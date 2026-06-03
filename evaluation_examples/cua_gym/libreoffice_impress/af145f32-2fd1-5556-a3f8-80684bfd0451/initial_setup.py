"""
Initial Setup: Create Investor_Data.xlsx with source data and open LibreOffice Impress
Task ID: impress_wf_049
Domain: libreoffice_impress
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'impress_wf_049'
DESKTOP = f'{WORKDIR}/Desktop'
OUTPUT = f'{DESKTOP}/Investor_Data.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    os.makedirs(DESKTOP, exist_ok=True)
    wb = openpyxl.Workbook()

    # --- Sheet 1: MRR ---
    ws_mrr = wb.active
    ws_mrr.title = 'MRR'
    ws_mrr.append(['Month', 'Value'])
    mrr_data = [
        ('Jan 2024', 245000),
        ('Feb 2024', 258000),
        ('Mar 2024', 271500),
        ('Apr 2024', 289000),
        ('May 2024', 305000),
        ('Jun 2024', 318000),
        ('Jul 2024', 334500),
        ('Aug 2024', 352000),
        ('Sep 2024', 371000),
        ('Oct 2024', 389000),
        ('Nov 2024', 408500),
        ('Dec 2024', 425000),
    ]
    for row in mrr_data:
        ws_mrr.append(list(row))

    # --- Sheet 2: Segments ---
    ws_seg = wb.create_sheet('Segments')
    ws_seg.append(['Segment', 'Revenue'])
    seg_data = [
        ('Enterprise', 1850000),
        ('Mid-Market', 1120000),
        ('SMB', 680000),
        ('Startup', 350000),
    ]
    for row in seg_data:
        ws_seg.append(list(row))

    # --- Sheet 3: Churn ---
    ws_churn = wb.create_sheet('Churn')
    ws_churn.append(['Month', 'Count', 'Rate'])
    churn_data = [
        ('Jan 2024', 12, 0.028),
        ('Feb 2024', 9, 0.021),
        ('Mar 2024', 15, 0.034),
        ('Apr 2024', 8, 0.018),
        ('May 2024', 11, 0.024),
        ('Jun 2024', 14, 0.030),
        ('Jul 2024', 7, 0.015),
        ('Aug 2024', 10, 0.021),
        ('Sep 2024', 13, 0.027),
    ]
    for row in churn_data:
        ws_churn.append(list(row))

    # --- Sheet 4: Financials ---
    ws_fin = wb.create_sheet('Financials')
    ws_fin.append(['Item', 'Q1', 'Q2', 'Q3'])
    fin_data = [
        ('Revenue', 774500, 912000, 1057500),
        ('COGS', 193625, 228000, 264375),
        ('Gross Profit', 580875, 684000, 793125),
        ('R&D', 232350, 273600, 317250),
        ('Sales & Marketing', 154900, 182400, 211500),
        ('G&A', 77450, 91200, 105750),
        ('EBITDA', 116175, 136800, 158625),
        ('Net Income', 85000, 102000, 121500),
    ]
    for row in fin_data:
        ws_fin.append(list(row))

    # --- Sheet 5: Risks ---
    ws_risks = wb.create_sheet('Risks')
    ws_risks.append(['Risk', 'Impact', 'Mitigation'])
    risk_data = [
        ('Enterprise churn spike in Q4', 'High', 'Dedicated CSM team for top 20 accounts; quarterly business reviews'),
        ('AWS cost escalation', 'Medium', 'Reserved instance optimization; multi-cloud evaluation in progress'),
        ('Key engineer attrition', 'High', 'Retention bonuses approved; career ladder restructuring underway'),
        ('Competitive pressure from Acme Corp', 'Medium', 'Accelerated feature roadmap; differentiation via AI capabilities'),
        ('Regulatory compliance (SOC 2 Type II)', 'Low', 'Audit preparation on track; expected completion by end of Q4'),
    ]
    for row in risk_data:
        ws_risks.append(list(row))

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Open LibreOffice Impress with a blank presentation
    launch_gui('libreoffice --impress', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Impress with DISPLAY=:0')


create_initial()
