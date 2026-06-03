"""
Initial Setup: Build environmental data source and open LibreOffice Impress
Task ID: impress_wf_058
Domain: libreoffice_impress
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
DESKTOP = f'{WORKDIR}/Desktop'
TASK_ID = 'impress_wf_058'
OUTPUT = f'{DESKTOP}/Env_Data.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    os.makedirs(DESKTOP, exist_ok=True)
    wb = openpyxl.Workbook()

    # --- Sheet 1: Emissions ---
    ws1 = wb.active
    ws1.title = 'Emissions'
    ws1.append(['Source', 'Tons'])
    emissions_data = [
        ['Transportation', 4250],
        ['Manufacturing', 8730],
        ['Office Operations', 1520],
        ['Supply Chain', 3680],
        ['Waste Disposal', 920],
        ['Agriculture', 2140],
    ]
    for row in emissions_data:
        ws1.append(row)

    # --- Sheet 2: Monthly ---
    ws2 = wb.create_sheet('Monthly')
    ws2.append(['Month', 'Actual', 'Target'])
    months = ['January', 'February', 'March', 'April', 'May', 'June',
              'July', 'August', 'September', 'October', 'November', 'December']
    actual_vals = [1850, 1780, 1720, 1690, 1650, 1580, 1540, 1510, 1470, 1430, 1390, 1350]
    target_vals = [1800, 1750, 1700, 1650, 1600, 1550, 1500, 1450, 1400, 1350, 1300, 1250]
    for i, m in enumerate(months):
        ws2.append([m, actual_vals[i], target_vals[i]])

    # --- Sheet 3: Energy ---
    ws3 = wb.create_sheet('Energy')
    ws3.append(['Month', 'Renewable', 'NonRenewable'])
    renewable = [320, 340, 360, 385, 410, 430, 455, 470, 490, 510, 530, 550]
    nonrenewable = [680, 660, 640, 615, 590, 570, 545, 530, 510, 490, 470, 450]
    for i, m in enumerate(months):
        ws3.append([m, renewable[i], nonrenewable[i]])

    # --- Sheet 4: Water ---
    ws4 = wb.create_sheet('Water')
    ws4.append(['Facility', 'Consumption'])
    water_data = [
        ['Headquarters', 45200],
        ['Manufacturing Plant A', 128500],
        ['Manufacturing Plant B', 97300],
        ['Distribution Center', 32800],
        ['Research Lab', 18600],
        ['Regional Office East', 12400],
    ]
    for row in water_data:
        ws4.append(row)

    # --- Sheet 5: Waste ---
    ws5 = wb.create_sheet('Waste')
    ws5.append(['Category', 'Current', 'Goal'])
    waste_data = [
        ['Paper & Cardboard', 72, 90],
        ['Plastics', 45, 75],
        ['Electronics', 58, 80],
        ['Organic Waste', 83, 95],
        ['Metal & Glass', 67, 85],
    ]
    for row in waste_data:
        ws5.append(row)

    # --- Sheet 6: Initiatives ---
    ws6 = wb.create_sheet('Initiatives')
    ws6.append(['Name', 'Status', 'Impact'])
    init_data = [
        ['Solar Panel Installation', 'Completed', 'High'],
        ['Fleet Electrification', 'In Progress', 'High'],
        ['Zero-Waste Cafeteria', 'Completed', 'Medium'],
        ['Rainwater Harvesting', 'Planning', 'Medium'],
        ['Carbon Offset Program', 'In Progress', 'High'],
        ['Green Building Certification', 'Completed', 'Low'],
    ]
    for row in init_data:
        ws6.append(row)

    # --- Sheet 7: YoY ---
    ws7 = wb.create_sheet('YoY')
    ws7.append(['Metric', '2022', '2023', 'Change'])
    yoy_data = [
        ['Carbon Emissions (tons)', 23400, 21240, '-9.2%'],
        ['Energy Usage (MWh)', 14200, 12800, '-9.9%'],
        ['Water Consumption (kL)', 358000, 334800, '-6.5%'],
        ['Waste Diverted (%)', 52, 65, '+25.0%'],
        ['Renewable Energy (%)', 32, 45, '+40.6%'],
        ['Green Certifications', 3, 5, '+66.7%'],
    ]
    for row in yoy_data:
        ws7.append(row)

    wb.save(OUTPUT)
    print(f'Initial data file created: {OUTPUT}')

    # Launch LibreOffice Impress (blank) for GUI-ready state
    launch_gui('libreoffice --impress', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Impress with DISPLAY=:0')


create_initial()
