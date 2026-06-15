"""
Initial Setup: Create Marketing_Analytics.xlsx with realistic marketing data
Task ID: impress_wf_079
Domain: libreoffice_impress
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
DESKTOP = f'{WORKDIR}/Desktop'
TASK_ID = 'impress_wf_079'
OUTPUT = f'{DESKTOP}/Marketing_Analytics.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    os.makedirs(DESKTOP, exist_ok=True)
    wb = openpyxl.Workbook()

    # --- Sheet 1: ChannelMix ---
    ws1 = wb.active
    ws1.title = 'ChannelMix'
    ws1.append(['Channel', 'Spend', 'Revenue'])
    channel_data = [
        ['Organic Search', 12500, 89200],
        ['Paid Search', 34800, 112500],
        ['Social Media', 18200, 54300],
        ['Email Marketing', 5600, 67800],
        ['Display Ads', 22100, 41600],
        ['Referral', 3200, 28900],
        ['Affiliate', 8900, 35400],
    ]
    for row in channel_data:
        ws1.append(row)

    # --- Sheet 2: WebAnalytics ---
    ws2 = wb.create_sheet('WebAnalytics')
    ws2.append(['Month', 'Sessions', 'Bounce', 'Duration'])
    web_data = [
        ['Jul-2025', 145200, 42.3, 3.45],
        ['Aug-2025', 158700, 39.8, 3.62],
        ['Sep-2025', 167400, 37.5, 3.78],
        ['Oct-2025', 152300, 41.1, 3.51],
        ['Nov-2025', 173800, 36.2, 3.89],
        ['Dec-2025', 189500, 34.7, 4.02],
    ]
    for row in web_data:
        ws2.append(row)

    # --- Sheet 3: SEO ---
    ws3 = wb.create_sheet('SEO')
    ws3.append(['Keyword', 'Rank', 'Change'])
    seo_data = [
        ['digital marketing tools', 3, 2],
        ['marketing automation software', 5, -1],
        ['email campaign platform', 2, 3],
        ['social media analytics', 8, -2],
        ['content marketing strategy', 4, 1],
        ['PPC management tool', 6, 4],
        ['SEO audit service', 1, 0],
        ['lead generation software', 7, -3],
        ['conversion rate optimization', 9, 2],
        ['marketing dashboard', 5, 1],
    ]
    for row in seo_data:
        ws3.append(row)

    # --- Sheet 4: PPC ---
    ws4 = wb.create_sheet('PPC')
    ws4.append(['Campaign', 'Cost', 'Clicks', 'Conversions'])
    ppc_data = [
        ['Brand Keywords', 4500, 12800, 640],
        ['Product Launch Q3', 8200, 18500, 555],
        ['Retargeting Display', 6100, 9200, 460],
        ['Competitor Terms', 7800, 14300, 429],
        ['Long-tail SEO', 3200, 8700, 522],
        ['Holiday Promo', 5600, 11400, 456],
    ]
    for row in ppc_data:
        ws4.append(row)

    # --- Sheet 5: Social ---
    ws5 = wb.create_sheet('Social')
    ws5.append(['Platform', 'Followers', 'Engagement'])
    social_data = [
        ['Instagram', 245000, 4.8],
        ['Facebook', 189000, 2.3],
        ['Twitter/X', 134000, 3.1],
        ['LinkedIn', 98000, 5.6],
        ['TikTok', 312000, 7.2],
        ['YouTube', 76000, 4.1],
    ]
    for row in social_data:
        ws5.append(row)

    # --- Sheet 6: Email ---
    ws6 = wb.create_sheet('Email')
    ws6.append(['Campaign', 'Sent', 'Opened', 'Clicked', 'Converted'])
    email_data = [
        ['Welcome Series', 45000, 18900, 5670, 1134],
        ['Product Update', 38000, 13300, 3990, 798],
        ['Monthly Newsletter', 52000, 19760, 4940, 988],
        ['Re-engagement', 28000, 7840, 2352, 470],
        ['Flash Sale Alert', 41000, 20500, 8200, 2460],
    ]
    for row in email_data:
        ws6.append(row)

    # --- Sheet 7: Content ---
    ws7 = wb.create_sheet('Content')
    ws7.append(['Title', 'Views', 'Shares', 'Conversions'])
    content_data = [
        ['Ultimate Guide to Marketing Automation', 24500, 1820, 245],
        ['10 SEO Trends for 2026', 31200, 2340, 312],
        ['How to Build a Content Calendar', 18700, 1120, 187],
        ['Social Media ROI Calculator', 42100, 3580, 421],
        ['Email Marketing Best Practices', 27800, 1950, 278],
        ['PPC Budget Optimization Guide', 15600, 890, 156],
        ['Customer Journey Mapping Tutorial', 21300, 1490, 213],
    ]
    for row in content_data:
        ws7.append(row)

    # --- Sheet 8: Attribution ---
    ws8 = wb.create_sheet('Attribution')
    ws8.append(['Channel', 'Percentage'])
    attribution_data = [
        ['Organic Search', 28],
        ['Paid Search', 22],
        ['Social Media', 15],
        ['Email', 18],
        ['Direct', 10],
        ['Referral', 7],
    ]
    for row in attribution_data:
        ws8.append(row)

    # --- Sheet 9: CAC ---
    ws9 = wb.create_sheet('CAC')
    ws9.append(['Month', 'Cost'])
    cac_data = [
        ['Jul-2025', 42.50],
        ['Aug-2025', 39.80],
        ['Sep-2025', 37.20],
        ['Oct-2025', 41.00],
        ['Nov-2025', 35.60],
        ['Dec-2025', 33.90],
    ]
    for row in cac_data:
        ws9.append(row)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Impress with a blank presentation
    launch_gui('libreoffice --impress', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Impress with DISPLAY=:0')


create_initial()
