"""
Initial Setup: Create Project_Data.xlsx with project data, open LibreOffice Impress
Task ID: impress_wf_043
Domain: libreoffice_impress
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
DESKTOP = f'{WORKDIR}/Desktop'
TASK_ID = 'impress_wf_043'
XLSX_OUTPUT = f'{DESKTOP}/Project_Data.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    os.makedirs(DESKTOP, exist_ok=True)

    wb = openpyxl.Workbook()

    # ---- Sheet 1: Milestones ----
    ws1 = wb.active
    ws1.title = 'Milestones'
    headers_ms = ['Milestone', 'Planned', 'Actual', 'Status']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    for col, h in enumerate(headers_ms, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')

    milestones_data = [
        ['Requirements Finalized', '2025-01-15', '2025-01-14', 'Completed'],
        ['Architecture Review', '2025-02-10', '2025-02-12', 'Completed'],
        ['Alpha Release', '2025-03-20', '2025-03-25', 'Delayed'],
        ['Beta Release', '2025-04-30', '2025-05-08', 'At Risk'],
        ['UAT Sign-off', '2025-06-01', '', 'Pending'],
        ['Production Go-Live', '2025-07-15', '', 'Pending'],
    ]
    for r, row_data in enumerate(milestones_data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    ws1.column_dimensions['A'].width = 28
    ws1.column_dimensions['B'].width = 14
    ws1.column_dimensions['C'].width = 14
    ws1.column_dimensions['D'].width = 14

    # ---- Sheet 2: Risks ----
    ws2 = wb.create_sheet('Risks')
    headers_risks = ['Risk', 'Probability', 'Impact', 'Mitigation']
    for col, h in enumerate(headers_risks, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')

    risks_data = [
        ['Key developer departure', 'Medium', 'High', 'Cross-training program; documentation sprint'],
        ['Third-party API deprecation', 'Low', 'Critical', 'Abstract integration layer; monitor vendor changelog'],
        ['Scope creep from stakeholders', 'High', 'Medium', 'Change advisory board; weekly scope review meetings'],
        ['Infrastructure cost overrun', 'Medium', 'Medium', 'Reserved instances; monthly cost audits'],
        ['Security vulnerability in deps', 'Medium', 'High', 'Automated Dependabot alerts; quarterly pen-test'],
    ]
    for r, row_data in enumerate(risks_data, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    ws2.column_dimensions['A'].width = 32
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 14
    ws2.column_dimensions['D'].width = 48

    # ---- Sheet 3: Resources ----
    ws3 = wb.create_sheet('Resources')
    headers_res = ['Team', 'Allocated', 'Used']
    for col, h in enumerate(headers_res, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')

    resources_data = [
        ['Frontend', 5, 4],
        ['Backend', 6, 7],
        ['QA', 3, 3],
        ['DevOps', 2, 2],
        ['Design', 2, 1],
    ]
    for r, row_data in enumerate(resources_data, 2):
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r, column=c, value=val)

    ws3.column_dimensions['A'].width = 16
    ws3.column_dimensions['B'].width = 14
    ws3.column_dimensions['C'].width = 14

    # ---- Sheet 4: Sprint ----
    ws4 = wb.create_sheet('Sprint')
    headers_sprint = ['Day', 'Planned', 'Actual']
    for col, h in enumerate(headers_sprint, 1):
        c = ws4.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')

    sprint_data = [
        [1, 50, 50],
        [2, 45, 47],
        [3, 40, 43],
        [4, 35, 39],
        [5, 30, 35],
        [6, 25, 30],
        [7, 20, 27],
        [8, 15, 22],
        [9, 10, 16],
        [10, 5, 10],
    ]
    for r, row_data in enumerate(sprint_data, 2):
        for c, val in enumerate(row_data, 1):
            ws4.cell(row=r, column=c, value=val)

    ws4.column_dimensions['A'].width = 10
    ws4.column_dimensions['B'].width = 14
    ws4.column_dimensions['C'].width = 14

    # ---- Sheet 5: Blockers ----
    ws5 = wb.create_sheet('Blockers')
    headers_block = ['Issue', 'Priority', 'Owner', 'Due']
    for col, h in enumerate(headers_block, 1):
        c = ws5.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')

    blockers_data = [
        ['CI/CD pipeline flaky tests', 'Critical', 'Amir Patel', '2025-04-10'],
        ['Staging environment out of disk', 'High', 'Lisa Nguyen', '2025-04-07'],
        ['OAuth token refresh race condition', 'High', 'James Rivera', '2025-04-12'],
        ['Missing accessibility audit report', 'Medium', 'Priya Sharma', '2025-04-15'],
    ]
    for r, row_data in enumerate(blockers_data, 2):
        for c, val in enumerate(row_data, 1):
            ws5.cell(row=r, column=c, value=val)

    ws5.column_dimensions['A'].width = 36
    ws5.column_dimensions['B'].width = 12
    ws5.column_dimensions['C'].width = 18
    ws5.column_dimensions['D'].width = 14

    wb.save(XLSX_OUTPUT)
    print(f'Initial file created: {XLSX_OUTPUT}')

    # Open LibreOffice Impress (blank) as the task starting point
    launch_gui('libreoffice --impress', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Impress with DISPLAY=:0')


create_initial()
