"""
Initial Setup: Create Board_Data.xlsx with financial data for board meeting presentation.
Task ID: impress_wf_089
Domain: libreoffice_impress
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'impress_wf_089'
DESKTOP = f'{WORKDIR}/Desktop'
OUTPUT_XLSX = f'{DESKTOP}/Board_Data.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    os.makedirs(DESKTOP, exist_ok=True)
    wb = openpyxl.Workbook()

    # --- Sheet 1: Financials ---
    ws1 = wb.active
    ws1.title = 'Financials'
    headers = ['Quarter', 'Revenue', 'EBITDA', 'NetIncome']
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
    data = [
        ['Q1 2024', 12500000, 3750000, 2100000],
        ['Q2 2024', 14200000, 4260000, 2450000],
        ['Q3 2024', 15800000, 5054000, 2900000],
        ['Q4 2023', 11800000, 3304000, 1850000],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # --- Sheet 2: PnL ---
    ws2 = wb.create_sheet('PnL')
    pnl_headers = ['Item', 'Q1', 'Q2', 'Q3', 'Budget']
    for c, h in enumerate(pnl_headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
    pnl_data = [
        ['Revenue', 12500000, 14200000, 15800000, 14000000],
        ['Cost of Goods Sold', 6250000, 6816000, 7110000, 7000000],
        ['Gross Profit', 6250000, 7384000, 8690000, 7000000],
        ['Operating Expenses', 2500000, 3124000, 3636000, 3200000],
        ['EBITDA', 3750000, 4260000, 5054000, 3800000],
        ['Depreciation & Amort.', 625000, 710000, 790000, 700000],
        ['Interest Expense', 312500, 355000, 395000, 350000],
        ['Tax Provision', 712500, 799750, 967250, 687500],
        ['Net Income', 2100000, 2450000, 2900000, 2062500],
    ]
    for r, row_data in enumerate(pnl_data, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    # --- Sheet 3: NPS ---
    ws3 = wb.create_sheet('NPS')
    nps_headers = ['Month', 'Score']
    for c, h in enumerate(nps_headers, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
    nps_data = [
        ['Jan 2024', 42], ['Feb 2024', 45], ['Mar 2024', 48],
        ['Apr 2024', 50], ['May 2024', 47], ['Jun 2024', 52],
        ['Jul 2024', 55], ['Aug 2024', 58], ['Sep 2024', 61],
    ]
    for r, row_data in enumerate(nps_data, 2):
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r, column=c, value=val)

    # --- Sheet 4: Churn ---
    ws4 = wb.create_sheet('Churn')
    churn_headers = ['Quarter', 'Rate']
    for c, h in enumerate(churn_headers, 1):
        cell = ws4.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
    churn_data = [
        ['Q1 2024', 4.2], ['Q2 2024', 3.8], ['Q3 2024', 3.1], ['Q4 2023', 5.0],
    ]
    for r, row_data in enumerate(churn_data, 2):
        for c, val in enumerate(row_data, 1):
            ws4.cell(row=r, column=c, value=val)

    # --- Sheet 5: Risks ---
    ws5 = wb.create_sheet('Risks')
    risk_headers = ['Risk', 'Severity', 'Mitigation', 'Owner']
    for c, h in enumerate(risk_headers, 1):
        cell = ws5.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
    risk_data = [
        ['Supply Chain Disruption', 'High', 'Dual sourcing strategy implemented', 'VP Operations - Sarah Kim'],
        ['Cybersecurity Breach', 'Critical', 'SOC2 audit completed; 24/7 monitoring', 'CISO - David Park'],
        ['Key Talent Attrition', 'Medium', 'Retention bonuses; career pathways', 'CHRO - Maria Santos'],
        ['Regulatory Compliance (GDPR)', 'High', 'External counsel engaged; DPO appointed', 'CLO - James Wright'],
        ['Currency Fluctuation Risk', 'Medium', 'Hedging program covering 80% exposure', 'CFO - Robert Chen'],
        ['Market Share Erosion', 'High', 'Accelerated product roadmap; pricing review', 'CMO - Lisa Zhang'],
    ]
    for r, row_data in enumerate(risk_data, 2):
        for c, val in enumerate(row_data, 1):
            ws5.cell(row=r, column=c, value=val)

    # --- Sheet 6: Pipeline ---
    ws6 = wb.create_sheet('Pipeline')
    pipe_headers = ['Target', 'Stage', 'Revenue', 'Valuation']
    for c, h in enumerate(pipe_headers, 1):
        cell = ws6.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
    pipe_data = [
        ['NovaTech Solutions', 'Due Diligence', 8500000, 42000000],
        ['DataStream Analytics', 'LOI Signed', 12000000, 72000000],
        ['CloudBridge Systems', 'Initial Screening', 5200000, 28000000],
        ['Vertex AI Labs', 'Negotiation', 15000000, 95000000],
    ]
    for r, row_data in enumerate(pipe_data, 2):
        for c, val in enumerate(row_data, 1):
            ws6.cell(row=r, column=c, value=val)

    wb.save(OUTPUT_XLSX)
    print(f'Initial file created: {OUTPUT_XLSX}')

    # Open LibreOffice Impress (blank, for the agent to create the presentation)
    launch_gui('libreoffice --impress', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Impress with DISPLAY=:0')


create_initial()
