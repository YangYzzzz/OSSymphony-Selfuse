"""
Initial Setup: Build ESG_Data.xlsx and open LibreOffice Impress
Task ID: impress_wf_092
Domain: libreoffice_impress
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
DESKTOP = f'{WORKDIR}/Desktop'
TASK_ID = 'impress_wf_092'
OUTPUT_XLSX = f'{DESKTOP}/ESG_Data.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_esg_data():
    os.makedirs(DESKTOP, exist_ok=True)
    wb = openpyxl.Workbook()

    # --- Sheet 1: Emissions ---
    ws = wb.active
    ws.title = 'Emissions'
    headers = ['Year', 'Scope1', 'Scope2', 'Scope3']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    emissions_data = [
        [2018, 45200, 32100, 128500],
        [2019, 42800, 30500, 125300],
        [2020, 38600, 27200, 112400],
        [2021, 36100, 25800, 105700],
        [2022, 33400, 23100, 98200],
        [2023, 30800, 21500, 91400],
    ]
    for r, row_data in enumerate(emissions_data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # --- Sheet 2: Energy ---
    ws2 = wb.create_sheet('Energy')
    for c, h in enumerate(['Source', 'Percentage'], 1):
        ws2.cell(row=1, column=c, value=h)
    energy_data = [
        ['Solar', 22],
        ['Wind', 18],
        ['Hydro', 8],
        ['Natural Gas', 30],
        ['Coal', 12],
        ['Nuclear', 10],
    ]
    for r, row_data in enumerate(energy_data, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    # --- Sheet 3: Water ---
    ws3 = wb.create_sheet('Water')
    for c, h in enumerate(['Facility', 'Usage', 'Target'], 1):
        ws3.cell(row=1, column=c, value=h)
    water_data = [
        ['HQ Campus', 125000, 100000],
        ['Manufacturing Plant A', 340000, 280000],
        ['Manufacturing Plant B', 285000, 230000],
        ['Data Center East', 95000, 75000],
        ['Data Center West', 88000, 70000],
        ['R&D Lab', 62000, 50000],
        ['Warehouse District', 45000, 38000],
    ]
    for r, row_data in enumerate(water_data, 2):
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r, column=c, value=val)

    # --- Sheet 4: Diversity ---
    ws4 = wb.create_sheet('Diversity')
    for c, h in enumerate(['Category', 'Male', 'Female', 'Other'], 1):
        ws4.cell(row=1, column=c, value=h)
    diversity_data = [
        ['Executive Leadership', 8, 5, 1],
        ['Senior Management', 22, 18, 2],
        ['Middle Management', 85, 72, 6],
        ['Technical Staff', 320, 195, 15],
        ['Operations', 410, 380, 22],
        ['Administrative', 45, 120, 8],
        ['Interns', 28, 35, 4],
    ]
    for r, row_data in enumerate(diversity_data, 2):
        for c, val in enumerate(row_data, 1):
            ws4.cell(row=r, column=c, value=val)

    # --- Sheet 5: Wellbeing ---
    ws5 = wb.create_sheet('Wellbeing')
    for c, h in enumerate(['Metric', 'Score'], 1):
        ws5.cell(row=1, column=c, value=h)
    wellbeing_data = [
        ['Work-Life Balance', 7.8],
        ['Job Satisfaction', 8.1],
        ['Mental Health Support', 7.2],
        ['Career Development', 7.5],
        ['Team Collaboration', 8.4],
        ['Management Support', 7.6],
        ['Workplace Safety', 9.1],
        ['Benefits Satisfaction', 7.9],
    ]
    for r, row_data in enumerate(wellbeing_data, 2):
        for c, val in enumerate(row_data, 1):
            ws5.cell(row=r, column=c, value=val)

    # --- Sheet 6: Community ---
    ws6 = wb.create_sheet('Community')
    for c, h in enumerate(['Initiative', 'Investment', 'Beneficiaries'], 1):
        ws6.cell(row=1, column=c, value=h)
    community_data = [
        ['STEM Education Program', 250000, 1200],
        ['Clean Water Initiative', 180000, 3500],
        ['Urban Reforestation', 120000, 8000],
        ['Digital Literacy Training', 95000, 650],
        ['Healthcare Access Fund', 310000, 4200],
        ['Youth Employment Program', 175000, 890],
    ]
    for r, row_data in enumerate(community_data, 2):
        for c, val in enumerate(row_data, 1):
            ws6.cell(row=r, column=c, value=val)

    # --- Sheet 7: Governance ---
    ws7 = wb.create_sheet('Governance')
    for c, h in enumerate(['Role', 'Name'], 1):
        ws7.cell(row=1, column=c, value=h)
    governance_data = [
        ['Board Chair', 'Dr. Elena Vasquez'],
        ['CEO', 'Michael Torres'],
        ['CFO', 'Priya Sharma'],
        ['COO', 'James Whitfield'],
        ['Chief Sustainability Officer', 'Amara Okafor'],
        ['VP Legal & Compliance', 'Robert Chang'],
        ['VP Human Resources', 'Sarah Mitchell'],
        ['Head of ESG Strategy', 'David Nakamura'],
    ]
    for r, row_data in enumerate(governance_data, 2):
        for c, val in enumerate(row_data, 1):
            ws7.cell(row=r, column=c, value=val)

    # --- Sheet 8: Compliance ---
    ws8 = wb.create_sheet('Compliance')
    for c, h in enumerate(['Policy', 'Status'], 1):
        ws8.cell(row=1, column=c, value=h)
    compliance_data = [
        ['Anti-Corruption Policy', 'Compliant'],
        ['Data Privacy (GDPR)', 'Compliant'],
        ['Environmental Management (ISO 14001)', 'Compliant'],
        ['Occupational Health & Safety', 'Compliant'],
        ['Supply Chain Due Diligence', 'In Progress'],
        ['Human Rights Policy', 'Compliant'],
        ['Whistleblower Protection', 'Compliant'],
        ['Carbon Disclosure (CDP)', 'Compliant'],
        ['Diversity & Inclusion Charter', 'In Progress'],
        ['Renewable Energy Commitment (RE100)', 'Compliant'],
    ]
    for r, row_data in enumerate(compliance_data, 2):
        for c, val in enumerate(row_data, 1):
            ws8.cell(row=r, column=c, value=val)

    wb.save(OUTPUT_XLSX)
    print(f'ESG_Data.xlsx created at: {OUTPUT_XLSX}')


create_esg_data()

# Open LibreOffice Impress (blank new presentation)
launch_gui('libreoffice --impress', delay_sec=2.0)
print('GUI_READY: LibreOffice Impress launched with DISPLAY=:0')
