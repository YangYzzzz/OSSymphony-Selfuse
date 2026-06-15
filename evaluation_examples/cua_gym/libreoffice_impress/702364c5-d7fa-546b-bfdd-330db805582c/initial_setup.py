"""
Initial Setup: Create Sales_Data.xlsx and open blank Impress presentation
Task ID: impress_wf_013
Domain: libreoffice_impress
"""

import os
import shlex
import subprocess
import time
import openpyxl
from datetime import date, timedelta

WORKDIR = '/home/user'
DESKTOP = f'{WORKDIR}/Desktop'
TASK_ID = 'impress_wf_013'
EXCEL_OUTPUT = f'{DESKTOP}/Sales_Data.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_sales_data():
    """Create Sales_Data.xlsx with 4 sheets of realistic Q3 data."""
    os.makedirs(DESKTOP, exist_ok=True)
    wb = openpyxl.Workbook()

    # --- Sheet 1: Monthly ---
    ws1 = wb.active
    ws1.title = 'Monthly'
    ws1.append(['Month', 'Revenue', 'Units'])
    monthly_data = [
        ['July', 128450, 1842],
        ['August', 145320, 2105],
        ['September', 137890, 1967],
    ]
    for row in monthly_data:
        ws1.append(row)

    # --- Sheet 2: Categories ---
    ws2 = wb.create_sheet('Categories')
    ws2.append(['Category', 'Percentage'])
    category_data = [
        ['Electronics', 35],
        ['Clothing', 25],
        ['Home & Garden', 18],
        ['Sports & Outdoors', 12],
        ['Books & Media', 10],
    ]
    for row in category_data:
        ws2.append(row)

    # --- Sheet 3: Daily ---
    ws3 = wb.create_sheet('Daily')
    ws3.append(['Date', 'Orders'])
    daily_orders = [
        62, 58, 71, 85, 93, 45, 42,
        68, 72, 79, 88, 95, 51, 47,
        73, 76, 82, 91, 98, 53, 49,
        78, 81, 87, 94, 102, 56, 52,
        83, 86,
    ]
    start_date = date(2025, 7, 1)
    for i, orders in enumerate(daily_orders):
        d = start_date + timedelta(days=i)
        ws3.append([d.strftime('%Y-%m-%d'), orders])

    # --- Sheet 4: Products ---
    ws4 = wb.create_sheet('Products')
    ws4.append(['Product', 'Revenue', 'Units', 'Rating'])
    products_data = [
        ['Wireless Earbuds Pro', 45230, 520, 4.7],
        ['Smart Watch Elite', 38750, 310, 4.5],
        ['Organic Cotton Tee', 28900, 1450, 4.3],
        ['Bamboo Desk Organizer', 22100, 880, 4.6],
        ['Running Shoes Ultra', 19850, 395, 4.4],
        ['Ceramic Plant Pot Set', 17600, 440, 4.2],
        ['LED Reading Lamp', 15320, 765, 4.1],
        ['Yoga Mat Premium', 12400, 620, 4.5],
        ['Stainless Water Bottle', 11200, 1120, 4.3],
        ['Bestseller Novel Bundle', 10450, 950, 4.0],
    ]
    for row in products_data:
        ws4.append(row)

    wb.save(EXCEL_OUTPUT)
    print(f'Sales_Data.xlsx created: {EXCEL_OUTPUT}')


def main():
    create_sales_data()

    # Open LibreOffice Impress with a blank presentation
    launch_gui('libreoffice --impress', delay_sec=3.0)
    print('GUI_READY: LibreOffice Impress launched with blank presentation')


main()
