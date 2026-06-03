"""
Initial Setup: Create annual_data.ods on Desktop with 4 sheets (Q1, Q2, Q3, Q4)
Task ID: osworld_multi_apps_doc_calc_to_writer_009
Domain: libreoffice_writer (multi-app: Calc source + Writer target)

Creates a LibreOffice Calc workbook with quarterly data.
Each sheet has a 5-row summary table (rows 1-6) and additional detail rows.
The agent task is to extract these summaries into an annual report in Writer.
"""

import os
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_doc_calc_to_writer_009'
# Place the source Calc file on the Desktop
DESKTOP = f'{WORKDIR}/Desktop'
OUTPUT = f'{DESKTOP}/annual_data.ods'
# .ods needs LibreOffice to create properly; we create .xlsx then convert
OUTPUT_XLSX = f'{WORKDIR}/annual_data_tmp.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    # Ensure Desktop directory exists
    os.makedirs(DESKTOP, exist_ok=True)
    # Ensure Documents directory exists (for agent to save output)
    os.makedirs(f'{WORKDIR}/Documents', exist_ok=True)

    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ #
    # Q1 Sheet
    # ------------------------------------------------------------------ #
    ws1 = wb.active
    ws1.title = 'Q1'

    header_font = Font(name='Calibri', bold=True, size=11)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Summary table header (row 1)
    q1_headers = ['Metric', 'Q1 Target ($)', 'Q1 Actual ($)', 'Variance (%)']
    for col, h in enumerate(q1_headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', bold=True, size=11, color='FFFFFFFF')
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Summary table data (rows 2-6) — 5 rows of quarterly summary metrics
    q1_summary = [
        ['Revenue',          1_250_000,  1_318_450,   5.48],
        ['Operating Costs',    820_000,    795_320,  -3.01],
        ['Gross Profit',       430_000,    523_130,  21.66],
        ['New Customers',        1_200,      1_347,  12.25],
        ['Customer Retention',    92.0,       94.3,   2.50],
    ]
    for r, row_data in enumerate(q1_summary, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.font = Font(name='Calibri', size=11)
            cell.border = border
            if c in (2, 3):
                cell.number_format = '#,##0'
            elif c == 4:
                cell.number_format = '0.00'

    # Detail data below summary (rows 8 onward) — more granular monthly breakdown
    ws1.cell(row=8, column=1, value='Monthly Breakdown')
    ws1.cell(row=8, column=1).font = Font(bold=True)

    monthly_headers = ['Month', 'Revenue', 'Costs', 'Headcount']
    for col, h in enumerate(monthly_headers, 1):
        ws1.cell(row=9, column=col, value=h).font = Font(bold=True)

    monthly_q1 = [
        ['January',  398_200, 261_100, 142],
        ['February', 432_700, 268_400, 143],
        ['March',    487_550, 265_820, 145],
    ]
    for r, row_data in enumerate(monthly_q1, 10):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Column widths
    for col_letter, width in zip('ABCD', [24, 16, 16, 14]):
        ws1.column_dimensions[col_letter].width = width

    # ------------------------------------------------------------------ #
    # Q2 Sheet
    # ------------------------------------------------------------------ #
    ws2 = wb.create_sheet('Q2')
    q2_headers = ['Metric', 'Q2 Target ($)', 'Q2 Actual ($)', 'Variance (%)']
    for col, h in enumerate(q2_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', bold=True, size=11, color='FFFFFFFF')
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    q2_summary = [
        ['Revenue',          1_380_000,  1_402_780,   1.65],
        ['Operating Costs',    895_000,    871_560,  -2.62],
        ['Gross Profit',       485_000,    531_220,   9.53],
        ['New Customers',        1_350,      1_289,  -4.52],
        ['Customer Retention',    93.0,       93.8,   0.86],
    ]
    for r, row_data in enumerate(q2_summary, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.font = Font(name='Calibri', size=11)
            cell.border = border
            if c in (2, 3):
                cell.number_format = '#,##0'
            elif c == 4:
                cell.number_format = '0.00'

    ws2.cell(row=8, column=1, value='Monthly Breakdown').font = Font(bold=True)
    for col, h in enumerate(['Month', 'Revenue', 'Costs', 'Headcount'], 1):
        ws2.cell(row=9, column=col, value=h).font = Font(bold=True)
    monthly_q2 = [
        ['April',  451_890, 289_600, 146],
        ['May',    468_320, 290_900, 148],
        ['June',   482_570, 291_060, 149],
    ]
    for r, row_data in enumerate(monthly_q2, 10):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)
    for col_letter, width in zip('ABCD', [24, 16, 16, 14]):
        ws2.column_dimensions[col_letter].width = width

    # ------------------------------------------------------------------ #
    # Q3 Sheet
    # ------------------------------------------------------------------ #
    ws3 = wb.create_sheet('Q3')
    q3_headers = ['Metric', 'Q3 Target ($)', 'Q3 Actual ($)', 'Variance (%)']
    for col, h in enumerate(q3_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', bold=True, size=11, color='FFFFFFFF')
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    q3_summary = [
        ['Revenue',          1_450_000,  1_511_340,   4.23],
        ['Operating Costs',    940_000,    918_720,  -2.26],
        ['Gross Profit',       510_000,    592_620,  16.20],
        ['New Customers',        1_400,      1_512,   8.00],
        ['Customer Retention',    94.0,       95.1,   1.17],
    ]
    for r, row_data in enumerate(q3_summary, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            cell.font = Font(name='Calibri', size=11)
            cell.border = border
            if c in (2, 3):
                cell.number_format = '#,##0'
            elif c == 4:
                cell.number_format = '0.00'

    ws3.cell(row=8, column=1, value='Monthly Breakdown').font = Font(bold=True)
    for col, h in enumerate(['Month', 'Revenue', 'Costs', 'Headcount'], 1):
        ws3.cell(row=9, column=col, value=h).font = Font(bold=True)
    monthly_q3 = [
        ['July',      491_240, 302_890, 150],
        ['August',    508_760, 307_400, 152],
        ['September', 511_340, 308_430, 153],
    ]
    for r, row_data in enumerate(monthly_q3, 10):
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r, column=c, value=val)
    for col_letter, width in zip('ABCD', [24, 16, 16, 14]):
        ws3.column_dimensions[col_letter].width = width

    # ------------------------------------------------------------------ #
    # Q4 Sheet
    # ------------------------------------------------------------------ #
    ws4 = wb.create_sheet('Q4')
    q4_headers = ['Metric', 'Q4 Target ($)', 'Q4 Actual ($)', 'Variance (%)']
    for col, h in enumerate(q4_headers, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', bold=True, size=11, color='FFFFFFFF')
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    q4_summary = [
        ['Revenue',          1_520_000,  1_489_670,  -1.99],
        ['Operating Costs',    985_000,    962_340,  -2.30],
        ['Gross Profit',       535_000,    527_330,  -1.43],
        ['New Customers',        1_450,      1_391,  -4.07],
        ['Customer Retention',    95.0,       94.7,  -0.32],
    ]
    for r, row_data in enumerate(q4_summary, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws4.cell(row=r, column=c, value=val)
            cell.font = Font(name='Calibri', size=11)
            cell.border = border
            if c in (2, 3):
                cell.number_format = '#,##0'
            elif c == 4:
                cell.number_format = '0.00'

    ws4.cell(row=8, column=1, value='Monthly Breakdown').font = Font(bold=True)
    for col, h in enumerate(['Month', 'Revenue', 'Costs', 'Headcount'], 1):
        ws4.cell(row=9, column=col, value=h).font = Font(bold=True)
    monthly_q4 = [
        ['October',   487_210, 318_760, 154],
        ['November',  498_320, 321_450, 155],
        ['December',  504_140, 322_130, 155],
    ]
    for r, row_data in enumerate(monthly_q4, 10):
        for c, val in enumerate(row_data, 1):
            ws4.cell(row=r, column=c, value=val)
    for col_letter, width in zip('ABCD', [24, 16, 16, 14]):
        ws4.column_dimensions[col_letter].width = width

    # Save as .xlsx first, then convert to .ods via LibreOffice headless
    wb.save(OUTPUT_XLSX)
    print(f'Temporary xlsx created: {OUTPUT_XLSX}')

    # Convert xlsx -> ods using LibreOffice headless
    env = os.environ.copy()
    env['DISPLAY'] = ':0'
    result = subprocess.run(
        ['libreoffice', '--headless', '--convert-to', 'ods',
         '--outdir', DESKTOP, OUTPUT_XLSX],
        capture_output=True, text=True, env=env, timeout=60
    )
    print('LibreOffice conversion stdout:', result.stdout)
    print('LibreOffice conversion stderr:', result.stderr)

    # The converted file will be named annual_data_tmp.ods in DESKTOP;
    # rename it to annual_data.ods
    converted = f'{DESKTOP}/annual_data_tmp.ods'
    if os.path.exists(converted):
        os.rename(converted, OUTPUT)
        print(f'Renamed to: {OUTPUT}')
    elif os.path.exists(OUTPUT):
        print(f'ODS file already at: {OUTPUT}')
    else:
        # Fallback: keep xlsx as .ods (will still work for the agent)
        import shutil
        shutil.copy(OUTPUT_XLSX, OUTPUT)
        print(f'Fallback: copied xlsx as ods to {OUTPUT}')

    # Cleanup temp xlsx
    if os.path.exists(OUTPUT_XLSX):
        os.remove(OUTPUT_XLSX)

    print(f'Annual data file created: {OUTPUT}')

    # GUI-ready: open the .ods file in LibreOffice Calc
    time.sleep(1)
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with annual_data.ods, DISPLAY=:0')


create_initial()
