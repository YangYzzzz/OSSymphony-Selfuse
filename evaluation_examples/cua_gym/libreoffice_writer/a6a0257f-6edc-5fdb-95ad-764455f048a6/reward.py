"""
Reward Script: Insert embedded Calc spreadsheet OLE object with sales data
Task ID: writer_fp_024
Domain: libreoffice_writer
Scoring:
  Component 1: OLE object exists in document (0.25)
  Component 2: OLE is an Excel/Calc spreadsheet type (0.15)
  Component 3: OLE object positioned after 'See table below:' text (0.15)
  Component 4: Embedded spreadsheet has correct structure - 5 rows, 3 cols, headers (0.20)
  Component 5: Embedded spreadsheet has correct Q1-Q4 data (0.25)
"""

import os
import zipfile
from io import BytesIO

WORKDIR = '/home/user'
TASK_ID = 'writer_fp_024'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load docx via python-docx for XML inspection
    try:
        from docx import Document
        doc = Document(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ns = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
    body = doc.element.body

    # Component 1: OLE object exists in document (0.25 points)
    try:
        ole_objects = body.findall('.//w:object', ns)
        ole_count = len(ole_objects)
        if ole_count > 0:
            print(f"PASS: Component 1 — Found {ole_count} OLE object(s) in document (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — No OLE objects found in document")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: OLE object is an Excel/Calc spreadsheet (0.15 points)
    try:
        ole_ns = {
            'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main',
            'o': 'urn:schemas-microsoft-com:office:office',
        }
        ole_obj_elements = body.findall('.//o:OLEObject', ole_ns)
        spreadsheet_prog_ids = [
            ole_elem.get('ProgID', '') for ole_elem in ole_obj_elements
            if any(kw in ole_elem.get('ProgID', '').lower() for kw in ['excel', 'calc', 'spreadsheet', 'opendocument.spreadsheet'])
        ]
        for ole_elem in ole_obj_elements:
            print(f"  OLE ProgID: {ole_elem.get('ProgID', '')}, Type: {ole_elem.get('Type', '')}")
        if len(spreadsheet_prog_ids) > 0:
            print(f"PASS: Component 2 — OLE object is a spreadsheet type (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 — OLE object ProgID does not indicate spreadsheet: {[e.get('ProgID', '') for e in ole_obj_elements]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: OLE object positioned after 'See table below:' (0.15 points)
    try:
        children = list(body)
        see_table_idx = None
        ole_para_idx = None
        for i, child in enumerate(children):
            tag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
            if tag == 'p':
                texts = child.findall('.//w:t', ns)
                text = ''.join(t.text or '' for t in texts).strip()
                if 'see table below' in text.lower():
                    see_table_idx = i
                has_ole = len(child.findall('.//w:object', ns)) > 0
                if has_ole and ole_para_idx is None:
                    ole_para_idx = i

        if see_table_idx is not None and ole_para_idx is not None:
            if ole_para_idx > see_table_idx:
                print(f"PASS: Component 3 — OLE object at body index {ole_para_idx} is after 'See table below:' at index {see_table_idx} (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 3 — OLE object at index {ole_para_idx} is NOT after 'See table below:' at index {see_table_idx}")
        elif ole_para_idx is None:
            print(f"FAIL: Component 3 — No paragraph containing OLE object found")
        else:
            print(f"FAIL: Component 3 — 'See table below:' text not found in document")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Components 4 & 5: Check embedded spreadsheet content
    # Extract embedded xlsx from the docx zip archive
    embedded_wb = None
    try:
        with zipfile.ZipFile(file_path, 'r') as z:
            # Find embedded spreadsheet files
            xlsx_files = [n for n in z.namelist() if n.endswith('.xlsx') and 'embed' in n.lower()]
            # Also check for .xls or .ods variants
            if not xlsx_files:
                xlsx_files = [n for n in z.namelist() if ('embed' in n.lower() or 'oleObject' in n) and (n.endswith('.xlsx') or n.endswith('.xls') or n.endswith('.ods'))]
            if xlsx_files:
                import openpyxl
                xlsx_data = z.read(xlsx_files[0])
                embedded_wb = openpyxl.load_workbook(BytesIO(xlsx_data))
                print(f"  Loaded embedded spreadsheet from: {xlsx_files[0]}")
            else:
                print(f"  No embedded xlsx found in archive. Files: {[n for n in z.namelist() if 'embed' in n.lower() or 'ole' in n.lower()]}")
    except Exception as e:
        print(f"  Could not extract embedded spreadsheet: {e}")

    # Component 4: Correct structure - 5 rows, 3 columns, correct headers (0.20 points)
    try:
        if embedded_wb is not None:
            ws = embedded_wb.active
            max_row = ws.max_row
            max_col = ws.max_column

            details = []
            checks_passed = 0
            total_checks = 4

            # Check dimensions: should be 5 rows, 3 columns
            if max_row >= 5 and max_col >= 3:
                details.append(f"dimensions {max_row}x{max_col} OK (>= 5x3)")
                checks_passed += 1
            else:
                details.append(f"dimensions {max_row}x{max_col} too small (need >= 5x3)")

            # Check headers in row 1
            h1 = str(ws.cell(1, 1).value or '').strip().lower()
            h2 = str(ws.cell(1, 2).value or '').strip().lower()
            h3 = str(ws.cell(1, 3).value or '').strip().lower()

            if 'quarter' in h1:
                details.append(f"header A1='{ws.cell(1,1).value}' OK")
                checks_passed += 1
            else:
                details.append(f"header A1='{ws.cell(1,1).value}' expected 'Quarter'")

            if 'revenue' in h2:
                details.append(f"header B1='{ws.cell(1,2).value}' OK")
                checks_passed += 1
            else:
                details.append(f"header B1='{ws.cell(1,2).value}' expected 'Revenue'")

            if 'growth' in h3:
                details.append(f"header C1='{ws.cell(1,3).value}' OK")
                checks_passed += 1
            else:
                details.append(f"header C1='{ws.cell(1,3).value}' expected 'Growth'")

            if checks_passed == total_checks:
                print(f"PASS: Component 4 — Correct structure: {'; '.join(details)} (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 4 — Structure issues: {'; '.join(details)}")
        else:
            print(f"FAIL: Component 4 — No embedded spreadsheet to check")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Correct Q1-Q4 sales data (0.25 points)
    try:
        if embedded_wb is not None:
            ws = embedded_wb.active
            # Expected data (case-insensitive, flexible matching)
            expected_data = {
                'q1': {'revenue': '1.2', 'growth': '5'},
                'q2': {'revenue': '1.4', 'growth': '8'},
                'q3': {'revenue': '1.1', 'growth': '-3'},
                'q4': {'revenue': '1.6', 'growth': '12'},
            }

            correct_rows = 0
            total_data_rows = 4

            for row_num in range(2, min(ws.max_row + 1, 7)):  # rows 2-6 max
                quarter_val = str(ws.cell(row_num, 1).value or '').strip().lower()
                revenue_val = str(ws.cell(row_num, 2).value or '').strip()
                growth_val = str(ws.cell(row_num, 3).value or '').strip()

                # Match quarter label
                quarter_key = None
                for qk in expected_data:
                    if qk in quarter_val:
                        quarter_key = qk
                        break

                if quarter_key is None:
                    print(f"  Row {row_num}: quarter '{quarter_val}' not recognized")
                    continue

                exp = expected_data[quarter_key]
                # Check revenue: look for the number (e.g., "1.2" in "$1.2M" or "1.2M" or "1200000")
                rev_match = exp['revenue'] in revenue_val.replace(',', '')
                # Check growth: look for the number (e.g., "5" in "5%" or "-3" in "-3%")
                growth_clean = growth_val.replace('%', '').strip()
                growth_match = growth_clean == exp['growth']

                if rev_match and growth_match:
                    correct_rows += 1
                    print(f"  Row {row_num}: {quarter_key.upper()} — revenue='{revenue_val}' growth='{growth_val}' CORRECT")
                else:
                    print(f"  Row {row_num}: {quarter_key.upper()} — revenue='{revenue_val}' (exp contains '{exp['revenue']}', match={rev_match}) growth='{growth_val}' (exp='{exp['growth']}', match={growth_match})")

            if correct_rows == total_data_rows:
                print(f"PASS: Component 5 — All {total_data_rows} quarters have correct data (0.25 pts)")
                total_score += 0.25
            elif correct_rows > 0:
                partial = round(0.25 * correct_rows / total_data_rows, 2)
                print(f"PARTIAL: Component 5 — {correct_rows}/{total_data_rows} quarters correct ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 5 — No quarters matched expected data")
        else:
            print(f"FAIL: Component 5 — No embedded spreadsheet to check")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.docx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
