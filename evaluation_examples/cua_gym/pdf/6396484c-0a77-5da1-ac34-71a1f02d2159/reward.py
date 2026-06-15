"""
FINAL REWARD SCRIPT - SUCCESS
Task: Extract the course schedule table from 'university_catalog.pdf' in /home/user/Documents/Academic (page 45) and save to 'course_list.xlsx'.
Generated: 2025-11-29 09:19:34
Status: success
Model: o3
Total Steps: 8
"""

from pathlib import Path
from openpyxl import load_workbook

"""
Reward script for:
Task: Extract the course schedule table from 'university_catalog.pdf' (page 45)
      and save it as 'course_list.xlsx' in /home/user/Documents/Academic.

The script verifies that:
1. The Excel file exists and can be opened.
2. There is exactly one worksheet and it has an appropriate title.
3. The header row matches the expected 5-column structure (Course Code, Course Name, Days, Time, Room).
4. At least five data rows are present.
5. All five expected example course records are present with correct names.

Each verification step contributes to a progressive score that sums to 1.0 only
when every requirement is met. The script prints detailed diagnostics for each
check and always prints the final score in the form "REWARD: X.X".
"""

EXPECTED_HEADER = ["Course Code", "Course Name", "Days", "Time", "Room"]
EXPECTED_COURSES = {
    "CS101": "Intro to Computer Science",
    "MATH201": "Calculus II",
    "PHYS150": "Physics I",
    "HIST210": "World History",
    "ENG102": "English Literature",
}
SHEET_NAME_CANDIDATES = ["Course Schedule", "Schedule", "Sheet1"]


def verify_course_schedule(xlsx_path: str) -> float:
    """Return a score between 0.0 and 1.0 based on task-completion level."""

    print(f"Verifying extracted course schedule at: {xlsx_path}\n")
    score = 0.0

    # -------------------------------------------------------------
    # 1) File existence & loading (no points for mere existence)
    # -------------------------------------------------------------
    path = Path(xlsx_path)
    if not path.exists():
        print("✗ Expected Excel file does not exist.")
        print("REWARD: 0.0")
        return 0.0

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        print(f"✗ Failed to load workbook: {e}")
        print("REWARD: 0.0")
        return 0.0

    # -------------------------------------------------------------
    # 2) Worksheet checks (0.10 total)
    # -------------------------------------------------------------
    sheet = wb.active

    if len(wb.sheetnames) == 1:
        print("✓ Exactly one worksheet present (0.05)")
        score += 0.05
    else:
        print(f"ℹ Multiple sheets detected: {wb.sheetnames} (0 points)")

    if any(sheet.title.lower() == name.lower() for name in SHEET_NAME_CANDIDATES):
        print(f"✓ Worksheet title '{sheet.title}' acceptable (0.05)")
        score += 0.05
    else:
        print(f"ℹ Unexpected worksheet title '{sheet.title}' (0 points)")

    # -------------------------------------------------------------
    # 3) Header validation (0.40 total)
    # -------------------------------------------------------------
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        print("✗ Worksheet is empty. No further checks possible.")
        print(f"REWARD: {score}")
        return score

    header = [str(c).strip() if c is not None else "" for c in rows[0]]

    # Count exact matches (case-insensitive) for the first five expected columns.
    matches = sum(
        1 for h, exp in zip(header, EXPECTED_HEADER) if h.lower() == exp.lower()
    )

    if matches == 5:
        print("✓ Header perfectly matches expected structure (0.30)")
        score += 0.30
    elif matches >= 4:
        print(f"ℹ Header partially matches ({matches}/5 correct) (0.15)")
        score += 0.15
    else:
        print(f"✗ Header mismatch. Found: {header} (0 points)")

    # Ensure there are at least five columns in total.
    if len(header) >= 5:
        print("✓ At least five columns present (0.10)")
        score += 0.10
    else:
        print(f"✗ Only {len(header)} columns found (0 points)")

    # -------------------------------------------------------------
    # 4) Data-row count check (0.10)
    # -------------------------------------------------------------
    data_rows = rows[1:]
    if len(data_rows) >= 5:
        print(f"✓ Contains {len(data_rows)} data rows (≥5) (0.10)")
        score += 0.10
    else:
        print(f"✗ Not enough data rows: {len(data_rows)} (0 points)")

    # -------------------------------------------------------------
    # 5) Expected course records verification (up to 0.40)
    # -------------------------------------------------------------
    found_courses = {
        str(r[0]).strip(): str(r[1]).strip()
        for r in data_rows
        if r and r[0] is not None and r[1] is not None
    }

    matched = sum(
        1
        for code, name in EXPECTED_COURSES.items()
        if code in found_courses and found_courses[code].lower() == name.lower()
    )
    print(f"Found {matched}/{len(EXPECTED_COURSES)} expected course rows")

    if matched == len(EXPECTED_COURSES):
        print("✓ All expected course rows present (0.40)")
        score += 0.40
    elif matched >= 3:
        print("ℹ Majority of expected courses present (0.20)")
        score += 0.20
    else:
        print("✗ Too few expected courses present (0 points)")

    # -------------------------------------------------------------
    # 6) Final score capping & output
    # -------------------------------------------------------------
    final_score = round(min(score, 1.0), 2)
    print(f"REWARD: {final_score}")
    return final_score


if __name__ == "__main__":
    verify_course_schedule("/home/user/Documents/Academic/course_list.xlsx")

