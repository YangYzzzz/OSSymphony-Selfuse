"""
Initial Setup: Create source files for PDF portfolio task
Task ID: pdf_gf1_048
Domain: pdf
"""

import os
import shlex
import subprocess
import time

WORKDIR = '/home/user'
TASK_ID = 'pdf_gf1_048'
FILES_DIR = f'{WORKDIR}/Documents/files'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    # Ensure directories exist
    os.makedirs(FILES_DIR, exist_ok=True)
    os.makedirs(f'{WORKDIR}/Documents', exist_ok=True)

    # --- 1. Create report.pdf (3 pages) ---
    import pymupdf

    doc = pymupdf.open()

    # Page 1: Title page
    page1 = doc.new_page(width=595, height=842)
    page1.insert_text(
        pymupdf.Point(150, 200),
        "Quarterly Performance Report",
        fontsize=24,
        fontname="hebo",
        color=(0.0, 0.1, 0.4),
    )
    page1.insert_text(
        pymupdf.Point(180, 260),
        "Meridian Analytics Group",
        fontsize=16,
        fontname="helv",
        color=(0.3, 0.3, 0.3),
    )
    page1.insert_text(
        pymupdf.Point(220, 310),
        "Q4 2025 — Confidential",
        fontsize=12,
        fontname="tiit",
        color=(0.4, 0.4, 0.4),
    )
    page1.insert_text(
        pymupdf.Point(200, 360),
        "Prepared by: Elena Vasquez, VP of Strategy",
        fontsize=11,
        fontname="helv",
        color=(0.3, 0.3, 0.3),
    )

    # Page 2: Executive Summary
    page2 = doc.new_page(width=595, height=842)
    page2.insert_text(
        pymupdf.Point(72, 72),
        "Executive Summary",
        fontsize=18,
        fontname="hebo",
        color=(0.0, 0.1, 0.4),
    )
    shape2 = page2.new_shape()
    shape2.draw_line(pymupdf.Point(72, 80), pymupdf.Point(523, 80))
    shape2.finish(color=(0.0, 0.1, 0.4), width=1.5)
    shape2.commit()

    summary_text = (
        "In Q4 2025, Meridian Analytics Group achieved record revenue of $14.7M, "
        "representing a 23% year-over-year increase. Our client retention rate improved "
        "to 94.2%, up from 89.1% in the previous quarter. The Enterprise Solutions division "
        "led growth with $6.3M in new contracts, driven primarily by demand for our predictive "
        "analytics platform.\n\n"
        "Key highlights include the successful launch of the RealTime Insights dashboard, "
        "onboarding of 18 new enterprise clients including Harmon Industries and Pacific "
        "Northwest Health Systems, and expansion of our data engineering team by 35%. "
        "Operating margins improved to 28.4%, exceeding our target of 25%.\n\n"
        "Looking ahead to Q1 2026, we project continued momentum with a pipeline of $9.2M "
        "in qualified opportunities. Strategic priorities include geographic expansion into "
        "the APAC region and the beta launch of our AI-powered anomaly detection module."
    )
    page2.insert_textbox(
        pymupdf.Rect(72, 100, 523, 700),
        summary_text,
        fontsize=11,
        fontname="helv",
        color=(0, 0, 0),
        align=pymupdf.TEXT_ALIGN_JUSTIFY,
    )

    # Page 3: Financial Highlights
    page3 = doc.new_page(width=595, height=842)
    page3.insert_text(
        pymupdf.Point(72, 72),
        "Financial Highlights",
        fontsize=18,
        fontname="hebo",
        color=(0.0, 0.1, 0.4),
    )
    shape3 = page3.new_shape()
    shape3.draw_line(pymupdf.Point(72, 80), pymupdf.Point(523, 80))
    shape3.finish(color=(0.0, 0.1, 0.4), width=1.5)
    shape3.commit()

    metrics = [
        ("Total Revenue:", "$14,723,450"),
        ("Operating Expenses:", "$10,539,800"),
        ("Net Income:", "$4,183,650"),
        ("Operating Margin:", "28.4%"),
        ("Client Retention:", "94.2%"),
        ("New Clients:", "18"),
        ("Headcount Growth:", "+35% (Data Engineering)"),
        ("Pipeline Q1 2026:", "$9,200,000"),
    ]
    y_pos = 110
    for label, value in metrics:
        page3.insert_text(pymupdf.Point(72, y_pos), label, fontsize=11, fontname="hebo", color=(0, 0, 0))
        page3.insert_text(pymupdf.Point(280, y_pos), value, fontsize=11, fontname="helv", color=(0.1, 0.3, 0.1))
        y_pos += 28

    doc.save(f'{FILES_DIR}/report.pdf')
    doc.close()
    print(f'Created: {FILES_DIR}/report.pdf')

    # --- 2. Create data.xlsx ---
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Q4 Revenue'

    headers = ['Month', 'Division', 'Revenue', 'Expenses', 'Net']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    data = [
        ['October', 'Enterprise Solutions', 2150000, 1420000, 730000],
        ['October', 'SMB Analytics', 980000, 710000, 270000],
        ['October', 'Consulting', 620000, 480000, 140000],
        ['November', 'Enterprise Solutions', 2280000, 1510000, 770000],
        ['November', 'SMB Analytics', 1050000, 740000, 310000],
        ['November', 'Consulting', 685000, 520000, 165000],
        ['December', 'Enterprise Solutions', 1870000, 1350000, 520000],
        ['December', 'SMB Analytics', 1120000, 790000, 330000],
        ['December', 'Consulting', 590000, 440000, 150000],
        ['October', 'Data Engineering', 780000, 580000, 200000],
        ['November', 'Data Engineering', 850000, 620000, 230000],
        ['December', 'Data Engineering', 1668450, 1359800, 308650],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    wb.save(f'{FILES_DIR}/data.xlsx')
    wb.close()
    print(f'Created: {FILES_DIR}/data.xlsx')

    # --- 3. Create readme.txt ---
    readme_content = """Meridian Analytics Group - Q4 2025 Data Package
================================================

This package contains the following files:

1. report.pdf
   - Quarterly Performance Report for Q4 2025
   - Includes executive summary and financial highlights
   - Prepared by Elena Vasquez, VP of Strategy

2. data.xlsx
   - Detailed revenue breakdown by month and division
   - Covers October through December 2025
   - Divisions: Enterprise Solutions, SMB Analytics, Consulting, Data Engineering

3. readme.txt (this file)
   - Package contents description and usage notes

CONFIDENTIALITY NOTICE:
This document package contains proprietary information belonging to
Meridian Analytics Group. Unauthorized distribution is prohibited.

For questions, contact: analytics-team@meridian-ag.com
Last updated: January 8, 2026
"""
    with open(f'{FILES_DIR}/readme.txt', 'w') as f:
        f.write(readme_content)
    print(f'Created: {FILES_DIR}/readme.txt')

    # Remove portfolio.pdf if it exists (ensure clean initial state)
    portfolio_path = f'{WORKDIR}/Documents/portfolio.pdf'
    if os.path.exists(portfolio_path):
        os.remove(portfolio_path)

    # --- GUI-ready startup ---
    # Open file manager showing the source files directory
    launch_gui(f'nautilus "{FILES_DIR}"', delay_sec=2.0)
    print('GUI_READY: launched nautilus with DISPLAY=:0')

create_initial()
