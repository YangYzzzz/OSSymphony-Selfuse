"""
Reward Script: Unlock cells A2:A10, lock cells B2:B10, and enable sheet protection
Task ID: calc_cop_protection_005
Domain: libreoffice_calc
Scoring:
  Component 1: Sheet protection is enabled (0.40 pts)
  Component 2: A2:A10 cells are unlocked (locked=False) (0.35 pts)
  Component 3: B2:B10 cells are locked (locked=True) AND VAT formulas intact (0.25 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_cop_protection_005'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: the sheet 'SimpleCalc' must exist
    if 'SimpleCalc' not in wb.sheetnames:
        print("FAIL: Sheet 'SimpleCalc' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['SimpleCalc']

    # Component 1: Sheet protection is enabled (0.40 points)
    # Task requires: "enable sheet protection with no password"
    # Initial state: protection.sheet == False
    # Golden state:  protection.sheet == True
    try:
        sheet_protected = ws.protection.sheet
        has_no_password = (ws.protection.password is None or ws.protection.password == '')
        if sheet_protected and has_no_password:
            print(f"PASS: Component 1 — Sheet protection enabled without password (0.40 pts)")
            total_score += 0.40
        elif sheet_protected and not has_no_password:
            print(f"FAIL: Component 1 — Sheet protection is enabled but has a password (unexpected)")
        else:
            print(f"FAIL: Component 1 — Sheet protection is NOT enabled (protection.sheet={sheet_protected})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: A2:A10 cells are unlocked (locked=False) (0.35 points)
    # Task requires: "Unlock cells A2:A10 so users can edit column A when protection is active"
    # Initial state: all cells have locked=True (default)
    # Golden state:  A2:A10 have locked=False
    try:
        unlocked_count = 0
        locked_wrongly = []
        for row in range(2, 11):  # rows 2 to 10 inclusive
            cell = ws.cell(row=row, column=1)  # column A
            if cell.protection.locked is False:
                unlocked_count += 1
            else:
                locked_wrongly.append(f"A{row}")

        if unlocked_count == 9:
            print(f"PASS: Component 2 — All 9 cells A2:A10 are unlocked (locked=False) (0.35 pts)")
            total_score += 0.35
        elif unlocked_count > 0:
            partial = round(0.35 * unlocked_count / 9, 4)
            print(f"PARTIAL: Component 2 — {unlocked_count}/9 cells in A2:A10 are unlocked; still locked: {locked_wrongly}")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — All A2:A10 cells are still locked; none have been unlocked")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: B2:B10 cells remain locked AND contain VAT formulas =An*1.21 (0.25 points)
    # Task requires: "lock cells B2:B10 (formula column B should not be editable)"
    # The cells were already locked=True in initial state, but we must confirm they remain so
    # and that the formulas are intact (no cell values or formulas should change)
    # This check FAILS on initial because initial has no sheet protection AND we combine
    # the formula integrity check with locked state — but since B cells were already locked
    # in the initial file, we scope this to verifying the complete combined requirement:
    # B cells locked=True AND formulas still valid AND sheet protection active.
    # Combined with Component 1 (sheet protection), this is only fully meaningful together.
    #
    # To ensure this component FAILS on initial: we gate this on sheet protection being active.
    # If sheet protection is off, locking B cells is meaningless — reward 0 for this component.
    try:
        locked_count = 0
        formula_count = 0
        issues = []
        for row in range(2, 11):  # rows 2 to 10
            cell = ws.cell(row=row, column=2)  # column B
            # Check cell is locked
            if cell.protection.locked is True:
                locked_count += 1
            else:
                issues.append(f"B{row} not locked (locked={cell.protection.locked})")
            # Check formula is intact: =A{row}*1.21
            expected_formula = f'=A{row}*1.21'
            actual_value = cell.value
            if isinstance(actual_value, str) and actual_value.upper().replace(' ', '') == expected_formula.upper().replace(' ', ''):
                formula_count += 1
            else:
                issues.append(f"B{row} formula incorrect: expected '{expected_formula}', found {repr(actual_value)}")

        # Only award points if sheet protection is enabled (makes the locking meaningful)
        # and B cells are locked with intact formulas
        if ws.protection.sheet and locked_count == 9 and formula_count == 9:
            print(f"PASS: Component 3 — All 9 cells B2:B10 locked with valid VAT formulas AND sheet protected (0.25 pts)")
            total_score += 0.25
        elif ws.protection.sheet and locked_count == 9 and formula_count < 9:
            partial = round(0.25 * formula_count / 9, 4)
            print(f"PARTIAL: Component 3 — B2:B10 locked but {9 - formula_count} formula(s) changed; issues: {issues}")
            total_score += partial
        elif not ws.protection.sheet:
            print(f"FAIL: Component 3 — Sheet protection not active, so locked B cells have no effect")
        else:
            print(f"FAIL: Component 3 — Issues found: {issues}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
