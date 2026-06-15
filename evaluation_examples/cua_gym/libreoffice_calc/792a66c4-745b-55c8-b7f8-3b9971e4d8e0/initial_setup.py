"""
Initial Setup: Spreadsheet with input values in A2:A10 and VAT formulas in B2:B10
Task ID: calc_cop_protection_005
Domain: libreoffice_calc

Creates a SimpleCalc sheet where:
- A2:A10 contain numeric input values (default lock state: Protected=True)
- B2:B10 contain formulas =A[n]*1.21 (VAT calculation, default lock state: Protected=True)
- Sheet protection is NOT enabled
- All cells use the default openpyxl lock (Protection(locked=True))
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Protection
from openpyxl.utils import get_column_letter

WORKDIR = '/home/user'
TASK_ID = 'calc_cop_protection_005'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: SimpleCalc ---
    ws = wb.active
    ws.title = 'SimpleCalc'

    # Headers
    ws['A1'] = 'Net Amount'
    ws['B1'] = 'Amount with VAT (21%)'

    # Style headers
    header_font = Font(name='Calibri', size=11, bold=True)
    ws['A1'].font = header_font
    ws['B1'].font = header_font

    # Realistic input values in A2:A10 (unit prices or quantities)
    input_values = [
        120.50,   # Office supplies
        349.00,   # Printer cartridge
        89.99,    # USB hub
        215.00,   # Monitor stand
        1250.00,  # Laptop sleeve
        45.75,    # Notebook pack
        175.00,   # Webcam
        530.00,   # Mechanical keyboard
        92.30,    # Cable organizer
    ]

    for row_idx, value in enumerate(input_values, start=2):
        # Column A: input value — default lock (Protected=True, the xlsx default)
        cell_a = ws.cell(row=row_idx, column=1, value=value)
        cell_a.number_format = '#,##0.00'
        # Explicitly set the default protection state (locked=True is openpyxl default)
        cell_a.protection = Protection(locked=True)

        # Column B: VAT formula — default lock (Protected=True)
        cell_b = ws.cell(row=row_idx, column=2, value=f'=A{row_idx}*1.21')
        cell_b.number_format = '#,##0.00'
        # Explicitly set the default protection state
        cell_b.protection = Protection(locked=True)

    # Column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 26

    # Sheet protection is NOT enabled in the initial file
    # (ws.protection.sheet remains False by default)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('  Sheet: SimpleCalc')
    print('  A2:A10: input values (locked=True by default, protection NOT enabled)')
    print('  B2:B10: formulas =A[n]*1.21 (locked=True by default, protection NOT enabled)')
    print('  Sheet protection: DISABLED')


create_initial()
