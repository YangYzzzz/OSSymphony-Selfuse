"""
Initial Setup: Apply email validation with @ check on C2
Task ID: calc_nrv_063
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_063'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Contacts"

    # --- Headers ---
    headers = ['Name', 'Phone', 'Email']
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Data rows (C2 intentionally left empty, rest have data) ---
    data = [
        ['Sarah Chen', '(415) 555-0123', None],  # C2 empty - this is the validation target
        ['Marcus Johnson', '(212) 555-0456', 'marcus.j@techcorp.com'],
        ['Priya Patel', '(650) 555-0789', 'priya.patel@globalinc.org'],
        ['James O\'Brien', '(310) 555-0234', 'jobrien@marketpro.net'],
        ['Wei Zhang', '(408) 555-0567', 'wei.zhang@datasync.io'],
        ['Olivia Martinez', '(773) 555-0890', 'olivia.m@brightpath.edu'],
        ['David Kim', '(617) 555-0345', 'dkim@innovatech.com'],
        ['Amara Osei', '(202) 555-0678', 'amara.osei@civicworks.gov'],
        ['Liam Murphy', '(503) 555-0912', 'liam.murphy@greenvale.co'],
        ['Fatima Al-Hassan', '(469) 555-0147', 'fatima.ah@medisolutions.com'],
        ['Carlos Rivera', '(305) 555-0258', 'crivera@suncoast.com'],
        ['Emily Watson', '(512) 555-0369', 'emily.w@buildwright.com'],
    ]

    data_font = Font(name="Calibri", size=11)
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = thin_border

    # --- Column widths ---
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 35

    # --- NO data validation on C2 (that is the task!) ---

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
