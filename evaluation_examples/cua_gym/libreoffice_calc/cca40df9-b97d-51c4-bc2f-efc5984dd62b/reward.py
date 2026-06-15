"""
FINAL REWARD SCRIPT - SUCCESS
Task: Please extract unique ISBN numbers from "ISBN with duplicates" to "Unique ISBNs" column. Keep the original sequence of first appearances. Finish without altering unrelated areas.
Generated: 2025-11-24 07:42:23
Status: success
Model: o3
Total Steps: 11
"""

import openpyxl
import glob
import os

def verify_unique_isbn_task():
    """Reward verifier for the ‘Unique ISBN extraction’ LibreOffice-Calc task.
    It returns a progressive score from 0.0 to 1.0 based on how accurately
    the user copied the first occurrence of every ISBN from the
    ‘ISBN with duplicates’ column into the ‘Unique ISBNs’ column, while
    leaving later duplicates blank and preserving other worksheet areas.
    """

    # ------------------------------------------------------------------
    # 1) Locate workbook (task gives a precise filename, but fall back to
    #    the first .xlsx in /home/user just in case).
    # ------------------------------------------------------------------
    default_path = (
        "/home/user/please_extract_unique_isbn_numbers_from_isbn_with_duplicates_to_unique_isbns_column_keep_the_origina.xlsx"
    )
    if os.path.exists(default_path):
        file_path = default_path
    else:
        xlsx_files = glob.glob("/home/user/*.xlsx")
        if not xlsx_files:
            print("✗ No .xlsx workbook found for verification")
            print("REWARD: 0.0")
            return 0.0
        file_path = xlsx_files[0]

    print(f"Workbook found: {file_path}")

    # ------------------------------------------------------------------
    # 2) Load workbook twice – once with formulas, once with cached values
    # ------------------------------------------------------------------
    try:
        wb_formula = openpyxl.load_workbook(file_path, data_only=False)
        wb_values  = openpyxl.load_workbook(file_path, data_only=True)
        sheet_f = wb_formula.active
        sheet_v = wb_values.active
        print("✓ Workbook successfully loaded")
    except Exception as exc:
        print(f"✗ Unable to load workbook: {exc}")
        print("REWARD: 0.0")
        return 0.0

    # ------------------------------------------------------------------
    # 3) Scoring weights (must sum to 1.0)
    # ------------------------------------------------------------------
    HEADER_WEIGHT          = 0.10  # headers unchanged
    STRUCTURE_WEIGHT       = 0.10  # expected columns still present
    UNIQUE_WEIGHT          = 0.55  # correct first-appearance ISBN copies
    DUPLICATE_WEIGHT       = 0.25  # later duplicates left blank

    total_score = 0.0

    # ------------------------------------------------------------------
    # 4) Header / structure checks
    # ------------------------------------------------------------------
    expected_headers = [
        "Title",
        "ISBN with duplicates",
        "Unique ISBNs",
        "Price",
    ]
    headers = [cell.value for cell in sheet_f[1]]

    if headers[: len(expected_headers)] == expected_headers:
        total_score += HEADER_WEIGHT
        print("✓ Headers unchanged")
    else:
        print(
            f"✗ Headers differ (found {headers[:len(expected_headers)]}, expected {expected_headers})"
        )

    if all(h in headers for h in expected_headers):
        total_score += STRUCTURE_WEIGHT
        print("✓ Worksheet structure intact (all expected columns present)")
    else:
        print("✗ Expected columns missing – structure altered")

    # Map key column indexes (positions identical between formula/value workbooks)
    try:
        dup_idx = headers.index("ISBN with duplicates")
        uniq_idx = headers.index("Unique ISBNs")
    except ValueError as e:
        print(f"✗ Required column not found: {e}")
        final = round(total_score, 2)
        print(f"REWARD: {final}")
        return final

    # ------------------------------------------------------------------
    # 5) Core ISBN extraction verification
    # ------------------------------------------------------------------
    seen_isbns = set()
    correct_unique = 0  # first occurrences correctly copied
    correct_blanks = 0  # duplicate rows correctly blank
    total_unique_expected = 0
    total_duplicate_expected = 0

    for r_idx, (row_f, row_v) in enumerate(
        zip(sheet_f.iter_rows(min_row=2), sheet_v.iter_rows(min_row=2, values_only=True)),
        start=2,
    ):
        isbn = row_v[dup_idx]
        cached_unique = row_v[uniq_idx]  # evaluated value (may be None)
        formula_or_value = row_f[uniq_idx].value  # raw (may be formula)

        is_first = isbn not in seen_isbns
        if is_first:
            total_unique_expected += 1
            # Accept either a literal match in cached value OR a formula (starts with '=')
            if cached_unique == isbn or (
                isinstance(formula_or_value, str) and formula_or_value.startswith("=")
            ):
                correct_unique += 1
            else:
                print(
                    f"✗ Row {r_idx}: expected unique ISBN {isbn}, found {cached_unique}"
                )
            seen_isbns.add(isbn)
        else:
            total_duplicate_expected += 1
            # Duplicate rows must be blank (None or "")
            if cached_unique is None or (
                isinstance(cached_unique, str) and cached_unique.strip() == ""
            ):
                correct_blanks += 1
            else:
                print(
                    f"✗ Row {r_idx}: duplicate ISBN should be blank, found {cached_unique}"
                )

    # ------------------------------------------------------------------
    # 6) Progressive scoring for core requirements
    # ------------------------------------------------------------------
    if total_unique_expected:
        total_score += UNIQUE_WEIGHT * (correct_unique / total_unique_expected)
    if total_duplicate_expected:
        total_score += DUPLICATE_WEIGHT * (correct_blanks / total_duplicate_expected)
    else:
        # Edge case: no duplicates – award full duplicate score
        total_score += DUPLICATE_WEIGHT

    final_score = round(min(total_score, 1.0), 2)

    # ------------------------------------------------------------------
    # 7) Summary & final output
    # ------------------------------------------------------------------
    print("--- Verification Summary ---")
    print(f"Correct unique cells     : {correct_unique}/{total_unique_expected}")
    print(f"Correct blank duplicates : {correct_blanks}/{total_duplicate_expected}")
    print(f"REWARD: {final_score}")
    return final_score

# Execute verification when script is run directly
if __name__ == "__main__":
    verify_unique_isbn_task()
