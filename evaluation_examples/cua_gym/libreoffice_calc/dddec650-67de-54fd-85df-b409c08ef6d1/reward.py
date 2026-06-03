"""
Reward Script: Inventory Reorder Alert System in LibreOffice Calc
Task ID: calc_grs_018
Domain: libreoffice_calc
Scoring:
  Component 1 — Order Status column with IF formulas (0.35 pts)
  Component 2 — COUNTIF summary formulas in row 2 (0.20 pts)
  Component 3 — Conditional formatting rules for ORDER NOW/SOON/OK (0.25 pts)
  Component 4 — Auto-filter on data range (0.10 pts)
  Component 5 — Merged header expanded to include column K (0.10 pts)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_018'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Inventory' sheet must exist
    if 'Inventory' not in wb.sheetnames:
        print("CRITICAL: 'Inventory' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Inventory']

    # Precondition: must have 25 data rows (rows 5-29) and header in row 4
    header_row4_col1 = ws.cell(row=4, column=1).value
    if header_row4_col1 != 'Product ID':
        print(f"CRITICAL: Expected 'Product ID' in A4, found: {header_row4_col1}")
        print("REWARD: 0.0")
        return 0.0

    # ===================================================================
    # Component 1: Order Status column with IF formulas (0.35 points)
    # This column does NOT exist in initial_env — only in golden_env
    # ===================================================================
    try:
        # Check header
        k_header = ws.cell(row=4, column=11).value
        if k_header and 'order' in str(k_header).lower() and 'status' in str(k_header).lower():
            # Check that at least 20 of 25 rows have the IF formula pattern
            formula_count = 0
            for r in range(5, 30):
                cell_val = ws.cell(row=r, column=11).value
                if cell_val and isinstance(cell_val, str) and cell_val.startswith('='):
                    upper_val = str(cell_val).upper()
                    # Must contain: IF, ORDER NOW, ORDER SOON, OK, and reference to stock vs minimum
                    has_if = 'IF(' in upper_val
                    has_order_now = 'ORDER NOW' in upper_val
                    has_order_soon = 'ORDER SOON' in upper_val
                    has_ok = '"OK"' in upper_val
                    # References current stock (D) and minimum stock (E) columns
                    has_stock_ref = ('D' + str(r)) in upper_val and ('E' + str(r)) in upper_val
                    if has_if and has_order_now and has_order_soon and has_ok and has_stock_ref:
                        formula_count += 1

            if formula_count >= 20:
                print(f"PASS: Component 1 — Order Status column has {formula_count}/25 valid IF formulas (0.35 pts)")
                total_score += 0.35
            elif formula_count >= 10:
                partial = 0.35 * (formula_count / 25)
                print(f"PARTIAL: Component 1 — Order Status column has {formula_count}/25 valid IF formulas ({partial:.2f} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 1 — Only {formula_count}/25 valid IF formulas in Order Status column")
        else:
            print(f"FAIL: Component 1 — 'Order Status' header not found in K4, found: {k_header}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ===================================================================
    # Component 2: COUNTIF summary formulas in row 2 (0.20 points)
    # Initial has labels but no formulas; golden has COUNTIF formulas
    # ===================================================================
    try:
        countif_found = 0
        # Expected: COUNTIF formulas counting "ORDER NOW", "ORDER SOON", "OK" in column K
        for col_idx in range(1, 12):
            cell_val = ws.cell(row=2, column=col_idx).value
            if cell_val and isinstance(cell_val, str) and '=COUNTIF' in cell_val.upper():
                upper_val = cell_val.upper().replace(' ', '')
                # Check it references K column range
                if 'K5:K29' in upper_val or 'K5:K' in upper_val or '$K' in upper_val:
                    countif_found += 1

        # Also check row 3 in case summary is placed there
        for col_idx in range(1, 12):
            cell_val = ws.cell(row=3, column=col_idx).value
            if cell_val and isinstance(cell_val, str) and '=COUNTIF' in cell_val.upper():
                upper_val = cell_val.upper().replace(' ', '')
                if 'K' in upper_val:
                    countif_found += 1

        if countif_found >= 3:
            print(f"PASS: Component 2 — Found {countif_found} COUNTIF summary formulas (0.20 pts)")
            total_score += 0.20
        elif countif_found >= 1:
            partial = 0.20 * (countif_found / 3)
            print(f"PARTIAL: Component 2 — Found {countif_found}/3 COUNTIF formulas ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No COUNTIF summary formulas found in rows 2-3")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ===================================================================
    # Component 3: Conditional formatting rules (0.25 points)
    # Initial has NO conditional formatting; golden has 3 rules
    # ===================================================================
    try:
        cf_rules = ws.conditional_formatting
        cf_count = 0
        has_order_now_rule = False
        has_order_soon_rule = False
        has_ok_rule = False

        for cf in cf_rules:
            for rule in cf.rules:
                formula_text = str(rule.formula).upper() if rule.formula else ''
                if 'ORDER NOW' in formula_text:
                    has_order_now_rule = True
                    cf_count += 1
                elif 'ORDER SOON' in formula_text:
                    has_order_soon_rule = True
                    cf_count += 1
                elif '"OK"' in formula_text:
                    has_ok_rule = True
                    cf_count += 1

        rules_found = sum([has_order_now_rule, has_order_soon_rule, has_ok_rule])
        if rules_found >= 3:
            print(f"PASS: Component 3 — All 3 conditional formatting rules present (ORDER NOW, ORDER SOON, OK) (0.25 pts)")
            total_score += 0.25
        elif rules_found >= 1:
            partial = 0.25 * (rules_found / 3)
            print(f"PARTIAL: Component 3 — {rules_found}/3 conditional formatting rules found ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No conditional formatting rules found for order statuses")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ===================================================================
    # Component 4: Auto-filter on data range (0.10 points)
    # Initial has NO auto-filter; golden has A4:K29
    # ===================================================================
    try:
        filter_ref = ws.auto_filter.ref
        if filter_ref:
            # Must include column K and cover the data area
            filter_upper = str(filter_ref).upper()
            # Check it covers at least columns A through K and includes row 4+
            if 'K' in filter_upper:
                print(f"PASS: Component 4 — Auto-filter set to {filter_ref} (0.10 pts)")
                total_score += 0.10
            else:
                print(f"PARTIAL: Component 4 — Auto-filter set to {filter_ref} but doesn't include column K (0.05 pts)")
                total_score += 0.05
        else:
            print(f"FAIL: Component 4 — No auto-filter defined")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ===================================================================
    # Component 5: Merged header expanded to include column K (0.10 points)
    # Initial merges A1:J1; golden merges A1:K1
    # ===================================================================
    try:
        merged_ranges = list(ws.merged_cells.ranges)
        header_merge_includes_k = False
        for mr in merged_ranges:
            mr_str = str(mr).upper()
            # Check if it's a row-1 merge that includes column K
            if mr_str.startswith('A1:') and 'K1' in mr_str:
                header_merge_includes_k = True
                break
            # Also accept wider merges
            if 'A1:' in mr_str and mr_str.endswith('1'):
                # Check if the end column is >= K (column 11)
                import re as re2
                match = re2.match(r'A1:([A-Z]+)1', mr_str)
                if match:
                    end_col = match.group(1)
                    if len(end_col) > 1 or end_col >= 'K':
                        header_merge_includes_k = True
                        break

        if header_merge_includes_k:
            print(f"PASS: Component 5 — Header merge includes column K (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 — Header merge does not extend to column K (ranges: {[str(mr) for mr in merged_ranges]})")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
