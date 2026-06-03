"""
Initial Setup: Inventory reorder alert system for restaurant supply chain
Task ID: calc_grs_018
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_018'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

CATEGORIES = ["Produce", "Dairy & Eggs", "Meat & Seafood", "Dry Goods", "Beverages"]

# 25 restaurant inventory items with realistic data
# Columns: Product ID, Product Name, Category, Current Stock, Min Stock, Max Stock, Reorder Qty, Lead Time (days), Unit Cost, Supplier
ITEMS = [
    ["PRD-001", "Roma Tomatoes (case)", "Produce", 8, 15, 60, 20, 2, 24.50, "FreshFarm Distributors"],
    ["PRD-002", "Romaine Lettuce (case)", "Produce", 12, 10, 40, 15, 2, 18.75, "FreshFarm Distributors"],
    ["PRD-003", "Yellow Onions (50lb)", "Produce", 3, 5, 25, 10, 2, 22.00, "Valley Growers Co-op"],
    ["PRD-004", "Fresh Basil (lb)", "Produce", 6, 4, 15, 8, 1, 12.50, "Herb Garden Supply"],
    ["PRD-005", "Red Bell Peppers (case)", "Produce", 5, 8, 30, 12, 3, 32.00, "FreshFarm Distributors"],
    ["PRD-006", "Heavy Cream (gal)", "Dairy & Eggs", 2, 6, 20, 8, 3, 8.95, "Meadowbrook Dairy"],
    ["PRD-007", "Large Eggs (case 15dz)", "Dairy & Eggs", 4, 3, 12, 6, 2, 42.00, "Meadowbrook Dairy"],
    ["PRD-008", "Parmesan Wheel (5lb)", "Dairy & Eggs", 1, 2, 8, 4, 5, 68.50, "Artisan Cheese Imports"],
    ["PRD-009", "Unsalted Butter (case)", "Dairy & Eggs", 7, 6, 20, 8, 3, 54.00, "Meadowbrook Dairy"],
    ["PRD-010", "Mozzarella (5lb block)", "Dairy & Eggs", 9, 5, 18, 6, 3, 28.75, "Artisan Cheese Imports"],
    ["PRD-011", "Chicken Breast (case 40lb)", "Meat & Seafood", 2, 4, 16, 6, 3, 89.00, "Premier Protein Co."],
    ["PRD-012", "Atlantic Salmon Fillet (lb)", "Meat & Seafood", 15, 10, 40, 15, 4, 14.50, "Ocean Fresh Seafood"],
    ["PRD-013", "Ground Beef 80/20 (10lb)", "Meat & Seafood", 3, 6, 24, 8, 3, 52.00, "Premier Protein Co."],
    ["PRD-014", "Pork Tenderloin (case)", "Meat & Seafood", 5, 4, 12, 4, 4, 76.00, "Premier Protein Co."],
    ["PRD-015", "Jumbo Shrimp 16/20 (5lb)", "Meat & Seafood", 4, 5, 20, 8, 5, 62.50, "Ocean Fresh Seafood"],
    ["PRD-016", "All-Purpose Flour (50lb)", "Dry Goods", 18, 10, 50, 20, 5, 19.50, "BulkPantry Wholesale"],
    ["PRD-017", "Extra Virgin Olive Oil (gal)", "Dry Goods", 3, 4, 16, 6, 4, 38.00, "Mediterranean Imports"],
    ["PRD-018", "Arborio Rice (25lb)", "Dry Goods", 12, 8, 30, 10, 5, 32.50, "BulkPantry Wholesale"],
    ["PRD-019", "San Marzano Tomatoes (case 6)", "Dry Goods", 6, 8, 36, 12, 5, 28.00, "Mediterranean Imports"],
    ["PRD-020", "Dried Penne Pasta (case 20lb)", "Dry Goods", 10, 8, 40, 15, 4, 24.00, "BulkPantry Wholesale"],
    ["PRD-021", "Sparkling Water (case 24)", "Beverages", 20, 15, 60, 20, 3, 16.50, "Metro Beverage Supply"],
    ["PRD-022", "House Red Wine (case 12)", "Beverages", 5, 6, 24, 8, 5, 96.00, "Vineyard Direct"],
    ["PRD-023", "Cold Brew Concentrate (gal)", "Beverages", 2, 3, 10, 4, 3, 28.00, "Roastworks Coffee"],
    ["PRD-024", "Fresh Orange Juice (gal)", "Beverages", 4, 5, 16, 6, 2, 11.50, "Metro Beverage Supply"],
    ["PRD-025", "Craft IPA (keg half-barrel)", "Beverages", 3, 2, 8, 3, 4, 145.00, "Hopside Brewing"],
]


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    # --- Summary section at top (labels only, NO formulas yet) ---
    header_font = Font(name="Calibri", size=14, bold=True)
    label_font = Font(name="Calibri", size=11, bold=True)

    ws.cell(row=1, column=1, value="Restaurant Supply Chain - Inventory Reorder Dashboard")
    ws["A1"].font = header_font
    ws.merge_cells("A1:J1")
    ws["A1"].alignment = Alignment(horizontal="center")

    # Summary labels row (row 2) - values intentionally left blank for the task
    ws.cell(row=2, column=1, value="Status Summary:")
    ws["A2"].font = label_font
    ws.cell(row=2, column=2, value="ORDER NOW:")
    ws["B2"].font = label_font
    # C2 intentionally empty - count formula goes here
    ws.cell(row=2, column=4, value="ORDER SOON:")
    ws["D2"].font = label_font
    # E2 intentionally empty - count formula goes here
    ws.cell(row=2, column=6, value="OK:")
    ws["F2"].font = label_font
    # G2 intentionally empty - count formula goes here

    # Row 3: blank separator

    # --- Data header row (row 4) ---
    headers = [
        "Product ID", "Product Name", "Category", "Current Stock",
        "Minimum Stock Level", "Maximum Stock Level", "Reorder Quantity",
        "Lead Time (days)", "Unit Cost", "Supplier"
    ]
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_font_white = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # --- Data rows (rows 5-29) ---
    for r, item in enumerate(ITEMS, 5):
        for c, val in enumerate(item, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 9:  # Unit Cost column
                cell.number_format = '$#,##0.00'
            if c in (4, 5, 6, 7, 8):  # numeric columns center-aligned
                cell.alignment = Alignment(horizontal="center")

    # --- Category dropdown validation ---
    dv = DataValidation(
        type="list",
        formula1='"' + ','.join(CATEGORIES) + '"',
        allow_blank=False,
        showDropDown=False,  # False = show the dropdown arrow
    )
    dv.error = "Please select a valid category"
    dv.errorTitle = "Invalid Category"
    dv.prompt = "Select product category"
    dv.promptTitle = "Category"
    dv.add("C5:C29")
    ws.add_data_validation(dv)

    # --- Column widths ---
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 26

    # Freeze header row
    ws.freeze_panes = "A5"

    # Row height for header
    ws.row_dimensions[4].height = 30

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
