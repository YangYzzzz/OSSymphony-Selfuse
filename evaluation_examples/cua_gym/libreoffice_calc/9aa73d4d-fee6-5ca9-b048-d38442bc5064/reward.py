"""
Reward Script: Build a delivery route efficiency tracker
Task ID: calc_ops_logistics_route_optimization_047
Domain: libreoffice_calc
Scoring:
  Component 1: K2:K51 formulas =En/Dn (stop completion rate) with % format (0.20 pts)
  Component 2: L2:L51 formulas =Fn/Gn (km efficiency) with % format (0.20 pts)
  Component 3: M2:M51 formulas =Jn/En (cost per stop) (0.20 pts)
  Component 4: Data validation dropdown on N2:N51 with correct options (0.20 pts)
  Component 5: Rows sorted by M ascending (cost per stop, most efficient first) (0.10 pts)
  Component 6: Conditional formatting on K column (red <80%, green >=90%) (0.10 pts)
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_ops_logistics_route_optimization_047'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet 'RouteTracker' exists
    if 'RouteTracker' not in wb.sheetnames:
        print("CRITICAL: Sheet 'RouteTracker' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['RouteTracker']

    # Component 1: K2:K51 contains stop completion rate formulas (=En/Dn) with % format (0.20 pts)
    # This FAILS on initial (K is empty) -> PASSES on golden (K has formulas with % format)
    try:
        k_formula_count = 0
        k_percent_count = 0
        for row in range(2, 52):
            cell = ws.cell(row=row, column=11)  # Column K
            val = cell.value
            if isinstance(val, str):
                # Accept formula like =E2/D2, =E10/D10, etc. (case-insensitive, possible spaces)
                normalized = val.upper().replace(' ', '')
                expected = f'=E{row}/D{row}'
                if normalized == expected:
                    k_formula_count += 1
            nf = cell.number_format or ''
            # Accept any percentage format: 0%, 0.00%, etc.
            if '%' in nf:
                k_percent_count += 1

        k_formula_ok = k_formula_count == 50
        k_percent_ok = k_percent_count >= 40  # allow some tolerance

        if k_formula_ok and k_percent_ok:
            print(f"PASS: Component 1 — K2:K51 all have =En/Dn formulas with % format ({k_formula_count}/50 formulas, {k_percent_count}/50 % format) (0.20 pts)")
            total_score += 0.20
        elif k_formula_ok:
            print(f"PARTIAL: Component 1 — K2:K51 formulas correct ({k_formula_count}/50) but % format missing ({k_percent_count}/50) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 — K2:K51 formulas incomplete ({k_formula_count}/50 correct formulas, {k_percent_count}/50 % format)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: L2:L51 contains km efficiency formulas (=Fn/Gn) with % format (0.20 pts)
    # This FAILS on initial (L is empty) -> PASSES on golden (L has formulas with % format)
    try:
        l_formula_count = 0
        l_percent_count = 0
        for row in range(2, 52):
            cell = ws.cell(row=row, column=12)  # Column L
            val = cell.value
            if isinstance(val, str):
                normalized = val.upper().replace(' ', '')
                expected = f'=F{row}/G{row}'
                if normalized == expected:
                    l_formula_count += 1
            nf = cell.number_format or ''
            if '%' in nf:
                l_percent_count += 1

        l_formula_ok = l_formula_count == 50
        l_percent_ok = l_percent_count >= 40

        if l_formula_ok and l_percent_ok:
            print(f"PASS: Component 2 — L2:L51 all have =Fn/Gn formulas with % format ({l_formula_count}/50 formulas, {l_percent_count}/50 % format) (0.20 pts)")
            total_score += 0.20
        elif l_formula_ok:
            print(f"PARTIAL: Component 2 — L2:L51 formulas correct ({l_formula_count}/50) but % format missing ({l_percent_count}/50) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 — L2:L51 formulas incomplete ({l_formula_count}/50 correct formulas, {l_percent_count}/50 % format)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: M2:M51 contains cost per stop formulas (=Jn/En) (0.20 pts)
    # This FAILS on initial (M is empty) -> PASSES on golden (M has formulas)
    try:
        m_formula_count = 0
        for row in range(2, 52):
            cell = ws.cell(row=row, column=13)  # Column M
            val = cell.value
            if isinstance(val, str):
                normalized = val.upper().replace(' ', '')
                expected = f'=J{row}/E{row}'
                if normalized == expected:
                    m_formula_count += 1

        if m_formula_count == 50:
            print(f"PASS: Component 3 — M2:M51 all have =Jn/En formulas ({m_formula_count}/50) (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 3 — M2:M51 formulas incomplete ({m_formula_count}/50 correct)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Data validation dropdown on N2:N51 with correct options (0.20 pts)
    # This FAILS on initial (no data validation on N) -> PASSES on golden
    try:
        dvs = ws.data_validations.dataValidation
        found_n_dv = False
        found_correct_options = False
        required_options = {'Completed', 'Partial', 'Cancelled', 'In Progress'}

        for dv in dvs:
            sqref_str = str(dv.sqref)
            # Check if this DV applies to N column (N2 or N2:N51 range)
            if 'N' in sqref_str and dv.type == 'list':
                found_n_dv = True
                formula = dv.formula1 or ''
                # Strip surrounding quotes
                formula_clean = formula.strip('"').strip("'")
                # Parse comma-separated options
                options = set(opt.strip() for opt in formula_clean.split(','))
                if options == required_options:
                    found_correct_options = True
                else:
                    print(f"  DV options found: {options}, required: {required_options}")

        if found_n_dv and found_correct_options:
            print(f"PASS: Component 4 — N2:N51 has correct dropdown (Completed, Partial, Cancelled, In Progress) (0.20 pts)")
            total_score += 0.20
        elif found_n_dv:
            print(f"FAIL: Component 4 — N column has data validation but options are incorrect")
        else:
            print(f"FAIL: Component 4 — No data validation found on N column")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Rows sorted by cost per stop (M) ascending (0.10 pts)
    # In the golden file, the data is reordered so that J/E is ascending
    # We compute M from raw J and E values and check ascending order
    # This FAILS on initial (initial has sequential route IDs, not sorted by M)
    # PASSES on golden (rows are sorted by J/E ascending)
    try:
        m_values = []
        valid_rows = 0
        for row in range(2, 52):
            j_val = ws.cell(row=row, column=10).value  # Fuel Cost
            e_val = ws.cell(row=row, column=5).value   # Stops Completed
            if j_val is not None and e_val is not None and e_val != 0:
                try:
                    m_computed = float(j_val) / float(e_val)
                    m_values.append(m_computed)
                    valid_rows += 1
                except (ValueError, TypeError):
                    pass

        if valid_rows >= 45:
            # Check if sorted ascending (allow minor floating point issues)
            is_sorted = all(m_values[i] <= m_values[i+1] + 0.0001 for i in range(len(m_values)-1))
            if is_sorted:
                print(f"PASS: Component 5 — Rows sorted by cost per stop (M) ascending ({valid_rows} valid rows checked) (0.10 pts)")
                total_score += 0.10
            else:
                # Find first violation
                violation_idx = next(i for i in range(len(m_values)-1) if m_values[i] > m_values[i+1] + 0.0001)
                print(f"FAIL: Component 5 — Rows not sorted by M ascending. First violation at data index {violation_idx}: {m_values[violation_idx]:.4f} > {m_values[violation_idx+1]:.4f}")
        else:
            print(f"FAIL: Component 5 — Not enough valid data rows ({valid_rows}) to check sort order")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Conditional formatting on K column (red fill if <80%, green fill if >=90%) (0.10 pts)
    # This FAILS on initial (no conditional formatting) -> PASSES on golden
    try:
        cf_rules_found = ws.conditional_formatting
        found_red_rule = False
        found_green_rule = False

        for cf_range in cf_rules_found:
            range_str = str(cf_range)
            # Check rules on K column
            if 'K' in range_str:
                for rule in cf_rules_found[cf_range]:
                    if rule.type == 'cellIs':
                        # Check for red fill rule (<80%)
                        if (rule.operator in ('lessThan',) and
                                rule.formula and '0.8' in str(rule.formula)):
                            if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                                fg = rule.dxf.fill.fgColor
                                if fg and hasattr(fg, 'rgb') and fg.rgb:
                                    # Red: FFFF0000 or similar red color
                                    rgb_val = fg.rgb.upper()
                                    if 'FF0000' in rgb_val or rgb_val == 'FFFF0000':
                                        found_red_rule = True
                        # Check for green fill rule (>=90%)
                        if (rule.operator in ('greaterThanOrEqual',) and
                                rule.formula and '0.9' in str(rule.formula)):
                            if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                                fg = rule.dxf.fill.fgColor
                                if fg and hasattr(fg, 'rgb') and fg.rgb:
                                    # Green: FF00FF00 or similar green color
                                    rgb_val = fg.rgb.upper()
                                    if '00FF00' in rgb_val or rgb_val == 'FF00FF00':
                                        found_green_rule = True

        if found_red_rule and found_green_rule:
            print(f"PASS: Component 6 — Conditional formatting on K: red (<80%) and green (>=90%) rules found (0.10 pts)")
            total_score += 0.10
        elif found_red_rule or found_green_rule:
            print(f"PARTIAL: Component 6 — Only partial conditional formatting: red={found_red_rule}, green={found_green_rule} (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 — No conditional formatting found on K column")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
