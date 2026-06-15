"""
Initial Setup: Delivery Route Efficiency Tracker
Task ID: calc_ops_logistics_route_optimization_047
Domain: libreoffice_calc

Creates a RouteTracker sheet with 50 route records (columns A-J filled).
Columns K, L, M are empty (formulas to be added by agent).
Column N has no dropdown (to be added by agent).
No sorting, no conditional formatting.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_logistics_route_optimization_047'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'RouteTracker'

    # --- Headers ---
    headers = [
        'Route ID',       # A
        'Driver',         # B
        'Date',           # C
        'Stops Planned',  # D
        'Stops Completed',# E
        'Planned km',     # F
        'Actual km',      # G
        'Planned Hours',  # H
        'Actual Hours',   # I
        'Fuel Cost',      # J
        'Stop Completion Rate',  # K — empty
        'km Efficiency',         # L — empty
        'Cost per Stop',         # M — empty
        'Route Status',          # N — needs dropdown
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.row_dimensions[1].height = 30

    # Column widths
    col_widths = {
        'A': 14, 'B': 18, 'C': 13, 'D': 14, 'E': 16,
        'F': 12, 'G': 12, 'H': 14, 'I': 13, 'J': 12,
        'K': 18, 'L': 14, 'M': 14, 'N': 14,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # --- Route Data (50 rows, realistic logistics data) ---
    drivers = [
        'Marco Rivera', 'Aisha Patel', 'Tom Nguyen', 'Lena Schmidt', 'James Okafor',
        'Sara Kim', 'Derek Walsh', 'Priya Sharma', 'Carlos Mendez', 'Nina Petrov',
        'Ethan Brooks', 'Fatima Hassan', 'Luke Andersson', 'Yuki Tanaka', 'Rosa Flores',
    ]

    from datetime import date, timedelta
    import random
    random.seed(42)

    base_date = date(2025, 1, 6)

    route_data = []
    for i in range(50):
        route_id = f'RT-{2025100 + i:07d}'
        driver = drivers[i % len(drivers)]
        route_date = (base_date + timedelta(days=i // 3)).strftime('%Y-%m-%d')
        stops_planned = random.randint(8, 22)
        # completion rate varies: some routes partial
        completion_pct = random.uniform(0.70, 1.0)
        stops_completed = max(1, round(stops_planned * completion_pct))
        planned_km = round(random.uniform(45, 180), 1)
        # actual km varies: can be more or less than planned
        actual_km = round(planned_km * random.uniform(0.80, 1.25), 1)
        planned_hours = round(planned_km / random.uniform(28, 38), 2)
        actual_hours = round(planned_hours * random.uniform(0.85, 1.30), 2)
        fuel_cost = round(actual_km * random.uniform(0.18, 0.32), 2)

        route_data.append([
            route_id, driver, route_date,
            stops_planned, stops_completed,
            planned_km, actual_km,
            planned_hours, actual_hours,
            fuel_cost,
            # K, L, M intentionally empty
            None, None, None,
            # N: no dropdown value yet, leave blank
            None,
        ])

    for r, row_vals in enumerate(route_data, 2):
        for c, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            # Format date column
            if c == 3 and val is not None:
                cell.number_format = 'yyyy-mm-dd'
            # Format numeric columns
            if c in (6, 7):
                cell.number_format = '0.0'
            if c in (8, 9):
                cell.number_format = '0.00'
            if c == 10:
                cell.number_format = '$#,##0.00'

    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: RouteTracker')
    print(f'  Rows: 51 (1 header + 50 data rows)')
    print(f'  Columns A-J filled, K-M empty, N empty (no dropdown)')


create_initial()
