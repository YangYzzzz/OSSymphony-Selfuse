"""
Initial Setup: Create a spreadsheet with task list and priority options for dropdown validation task.
Task ID: calc_nrv_075
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.workbook.defined_name import DefinedName

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_075'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Headers
    ws["A1"] = "Task"
    ws["B1"] = "Priority"

    # Realistic task data in A2:A15 (B2:B15 left empty - no validation yet)
    tasks = [
        "Review quarterly budget report",
        "Update employee onboarding docs",
        "Fix login page timeout issue",
        "Prepare client presentation slides",
        "Conduct team performance reviews",
        "Migrate database to new server",
        "Design new landing page mockup",
        "Write API documentation for v2.0",
        "Schedule vendor contract renewal",
        "Test payment gateway integration",
        "Organize company retreat logistics",
        "Audit security access permissions",
        "Create marketing campaign timeline",
        "Optimize search query performance",
    ]
    for r, task in enumerate(tasks, 2):
        ws.cell(row=r, column=1, value=task)

    # Priority options in F1:F4
    ws["F1"] = "Critical"
    ws["F2"] = "High"
    ws["F3"] = "Medium"
    ws["F4"] = "Low"

    # Create named range 'Priorities' referring to Sheet1!$F$1:$F$4
    dn = DefinedName("Priorities", attr_text="Sheet1!$F$1:$F$4")
    wb.defined_names.add(dn)

    # B2:B15 are empty, NO validation applied

    wb.save(OUTPUT)
    print(f"Initial file created: {OUTPUT}")

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print("GUI_READY: launched LibreOffice Calc with DISPLAY=:0")


create_initial()
