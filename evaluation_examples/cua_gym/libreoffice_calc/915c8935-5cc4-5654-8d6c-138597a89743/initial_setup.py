"""
Initial Setup: Calculate the annualized attrition rate and cost of turnover for each department.
Task ID: calc_hr_047
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_047'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Attrition ---
    ws = wb.active
    ws.title = 'Attrition'

    # Headers
    headers = ['Department', 'Avg Headcount', 'Departures (YTD)', 'Avg Salary',
               'Attrition Rate', 'Replacement Cost (1.5x salary)']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_side = Side(style='thin', color='000000')
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Data rows (columns A-D only; E and F left empty for the task)
    data = [
        ['Engineering', 50, 8, 105000],
        ['Sales', 35, 12, 72000],
        ['HR', 15, 2, 68000],
        ['Marketing', 20, 5, 75000],
    ]

    data_font = Font(name='Calibri', size=11)
    data_align = Alignment(horizontal='center', vertical='center')

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = border
        # Also apply border and alignment to empty E and F cells
        for c in [5, 6]:
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = data_align
            cell.font = data_font

    # Format D column as currency
    for r in range(2, 6):
        ws.cell(row=r, column=4).number_format = '$#,##0'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 28

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
