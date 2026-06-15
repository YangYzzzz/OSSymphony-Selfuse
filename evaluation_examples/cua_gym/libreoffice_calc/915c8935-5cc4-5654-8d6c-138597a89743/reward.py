"""
Reward Script: Calculate annualized attrition rate and cost of turnover for each department
Task ID: calc_hr_047
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): E2:E5 contain attrition rate formulas (=C/B)
  Component 2 (0.30): F2:F5 contain replacement cost formulas (=C*D*1.5)
  Component 3 (0.20): E column formatted as percentage
  Component 4 (0.20): F column formatted as currency/number
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_047'


def normalize_formula(f):
    """Normalize formula for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet exists
    if 'Attrition' not in wb.sheetnames:
        print("CRITICAL: 'Attrition' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Attrition']

    # Expected attrition rate formulas: E = C/B
    expected_e = {
        2: '=C2/B2',
        3: '=C3/B3',
        4: '=C4/B4',
        5: '=C5/B5',
    }

    # Expected replacement cost formulas: F = C * D * 1.5
    expected_f = {
        2: '=C2*D2*1.5',
        3: '=C3*D3*1.5',
        4: '=C4*D4*1.5',
        5: '=C5*D5*1.5',
    }

    # Component 1: E2:E5 attrition rate formulas (0.30 points)
    # Each correct formula earns 0.075 points
    try:
        e_score = 0.0
        for row, expected in expected_e.items():
            cell_val = ws.cell(row=row, column=5).value
            norm_actual = normalize_formula(cell_val)
            norm_expected = normalize_formula(expected)
            if norm_actual == norm_expected:
                print(f"PASS: E{row} formula correct: {cell_val}")
                e_score += 0.075
            else:
                # Also accept equivalent forms like =C2/B2 vs formulas that compute the same
                # Check if the value is a number close to C/B (agent may have entered value directly)
                c_val = ws.cell(row=row, column=3).value
                b_val = ws.cell(row=row, column=2).value
                if isinstance(cell_val, (int, float)) and b_val and c_val:
                    expected_val = c_val / b_val
                    if abs(cell_val - expected_val) < 0.001:
                        print(f"PASS: E{row} has correct computed value: {cell_val} (expected ~{expected_val:.4f})")
                        e_score += 0.075
                    else:
                        print(f"FAIL: E{row} expected formula {expected} or value ~{expected_val:.4f}, found: {cell_val!r}")
                else:
                    print(f"FAIL: E{row} expected formula {expected}, found: {cell_val!r}")
        if e_score > 0:
            print(f"Component 1 subtotal: {e_score:.3f}/0.300")
            total_score += e_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: F2:F5 replacement cost formulas (0.30 points)
    # Each correct formula earns 0.075 points
    try:
        f_score = 0.0
        for row, expected in expected_f.items():
            cell_val = ws.cell(row=row, column=6).value
            norm_actual = normalize_formula(cell_val)
            norm_expected = normalize_formula(expected)
            if norm_actual == norm_expected:
                print(f"PASS: F{row} formula correct: {cell_val}")
                f_score += 0.075
            else:
                # Accept equivalent computed values
                c_val = ws.cell(row=row, column=3).value
                d_val = ws.cell(row=row, column=4).value
                if isinstance(cell_val, (int, float)) and c_val and d_val:
                    expected_val = c_val * d_val * 1.5
                    if abs(cell_val - expected_val) < 1.0:
                        print(f"PASS: F{row} has correct computed value: {cell_val} (expected {expected_val})")
                        f_score += 0.075
                    else:
                        print(f"FAIL: F{row} expected formula {expected} or value {expected_val}, found: {cell_val!r}")
                else:
                    print(f"FAIL: F{row} expected formula {expected}, found: {cell_val!r}")
        if f_score > 0:
            print(f"Component 2 subtotal: {f_score:.3f}/0.300")
            total_score += f_score
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: E column percentage format (0.20 points)
    # Check that E2:E5 have percentage number format
    try:
        pct_count = 0
        for row in range(2, 6):
            nf = ws.cell(row=row, column=5).number_format
            if '%' in str(nf):
                pct_count += 1
                print(f"PASS: E{row} number_format contains '%': {nf}")
            else:
                print(f"FAIL: E{row} number_format should contain '%', found: {nf!r}")
        if pct_count == 4:
            print(f"PASS: Component 3 — all 4 E cells formatted as percentage (0.20 pts)")
            total_score += 0.20
        elif pct_count > 0:
            partial = 0.20 * (pct_count / 4)
            print(f"PARTIAL: Component 3 — {pct_count}/4 E cells formatted as percentage ({partial:.2f} pts)")
            total_score += partial
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: F column currency/number format (0.20 points)
    # Check that F2:F5 have a currency or number format (not General)
    try:
        fmt_count = 0
        for row in range(2, 6):
            nf = ws.cell(row=row, column=6).number_format
            # Accept any non-General format that contains $ or # or 0 (numeric formatting)
            if nf and nf != 'General' and ('$' in str(nf) or '#' in str(nf) or '0' in str(nf)):
                fmt_count += 1
                print(f"PASS: F{row} number_format is formatted: {nf}")
            else:
                print(f"FAIL: F{row} number_format should be currency/number, found: {nf!r}")
        if fmt_count == 4:
            print(f"PASS: Component 4 — all 4 F cells formatted as currency/number (0.20 pts)")
            total_score += 0.20
        elif fmt_count > 0:
            partial = 0.20 * (fmt_count / 4)
            print(f"PARTIAL: Component 4 — {fmt_count}/4 F cells formatted ({partial:.2f} pts)")
            total_score += partial
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
