"""
Initial Setup: GPA grade tracker with letter grades in column A, empty grade points in column B
Task ID: calc_fma_switch_grade_077
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_fma_switch_grade_077'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: GPA ---
    ws = wb.active
    ws.title = 'GPA'

    # Headers
    ws['A1'] = 'Grade'
    ws['B1'] = 'Grade Points'
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)

    # Column A data: letter grades (rows 2-16)
    # Sequence: A, B, C, A, F, B, D, C, A, B, F, C, D, A, B
    grades = ['A', 'B', 'C', 'A', 'F', 'B', 'D', 'C', 'A', 'B', 'F', 'C', 'D', 'A', 'B']
    for i, grade in enumerate(grades, start=2):
        ws.cell(row=i, column=1, value=grade)

    # Column B (B2:B16): EMPTY — task requires adding SWITCH formulas here

    # Set column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: GPA')
    print(f'  Column A (A2:A16): Letter grades (A, B, C, A, F, B, D, C, A, B, F, C, D, A, B)')
    print(f'  Column B (B2:B16): Empty (ready for SWITCH formulas)')


create_initial()
