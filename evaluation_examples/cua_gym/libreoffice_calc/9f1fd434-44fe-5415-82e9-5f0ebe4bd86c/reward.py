"""
Reward Script: Use SWITCH to map letter grades to grade point values in column B
Task ID: calc_fma_switch_grade_077
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.5): All 15 cells B2:B16 contain a SWITCH formula
  - Component 2 (0.3): SWITCH formula has correct grade-to-points mapping (A=4.0, B=3.0, C=2.0, D=1.0, F=0.0)
  - Component 3 (0.2): Compound check — formulas present AND Column A data / B1 header are intact
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fma_switch_grade_077'


def check_formula_has_correct_mappings(formula_str):
    """
    Returns the number of correct grade-to-point mappings found in the SWITCH formula.
    Expected mappings: A=4.0, B=3.0, C=2.0, D=1.0, F=0.0
    """
    required_mappings = [
        ('"A"', '4'),    # A=4.0
        ('"B"', '3'),    # B=3.0
        ('"C"', '2'),    # C=2.0
        ('"D"', '1'),    # D=1.0
        ('"F"', '0'),    # F=0.0
    ]
    formula_normalized = formula_str.replace(' ', '')
    found_count = sum(
        1 for grade_str, point_str in required_mappings
        if re.search(re.escape(grade_str) + r',' + re.escape(point_str), formula_normalized, re.IGNORECASE)
    )
    return found_count  # returns 0-5; all 5 = correct mapping


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the GPA sheet exists (precondition gate)
    if 'GPA' not in wb.sheetnames:
        print("CRITICAL: 'GPA' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['GPA']

    # Component 1: All 15 cells B2:B16 contain a SWITCH formula (0.5 points)
    # This FAILS on initial (B2:B16 are None) and PASSES on golden (SWITCH formulas present)
    try:
        switch_count = 0
        non_switch_cells = []
        for row in range(2, 17):  # rows 2-16 inclusive (15 cells)
            cell_val = ws.cell(row=row, column=2).value
            if cell_val is not None and isinstance(cell_val, str) and cell_val.upper().startswith('=SWITCH('):
                switch_count += 1
            else:
                non_switch_cells.append(f"B{row}={repr(cell_val)}")

        if switch_count == 15:
            print(f"PASS: Component 1 — All 15 cells B2:B16 contain SWITCH formulas (0.5 pts)")
            total_score += 0.5
        elif switch_count > 0:
            # Partial credit: some cells have SWITCH formulas (proportional)
            partial_c1 = round(0.5 * switch_count / 15, 4)
            print(f"PARTIAL: Component 1 — {switch_count}/15 cells have SWITCH formulas (partial {partial_c1} pts)")
            print(f"  Missing/wrong cells: {non_switch_cells[:5]}")
            if partial_c1 > 0:
                total_score += partial_c1
        else:
            print(f"FAIL: Component 1 — No SWITCH formulas found in B2:B16")
            print(f"  Sample cells: {non_switch_cells[:5]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: SWITCH formula correctness — maps A=4.0, B=3.0, C=2.0, D=1.0, F=0.0 (0.3 points)
    # Checks that each SWITCH formula contains the correct grade-to-point mappings.
    # This FAILS on initial (no formulas) and PASSES on golden (correct mappings).
    try:
        correct_formulas = 0
        incorrect_cells = []

        for row in range(2, 17):
            cell_val = ws.cell(row=row, column=2).value
            if cell_val is None or not isinstance(cell_val, str):
                incorrect_cells.append(f"B{row}: empty/non-formula")
                continue

            if not cell_val.upper().startswith('=SWITCH('):
                incorrect_cells.append(f"B{row}: not a SWITCH formula")
                continue

            # Count correct mappings in this formula (need all 5 to count as correct)
            mapping_count = check_formula_has_correct_mappings(cell_val)
            if mapping_count == 5:
                correct_formulas += 1
            else:
                incorrect_cells.append(f"B{row}: only {mapping_count}/5 mappings correct in {repr(cell_val[:50])}")

        if correct_formulas == 15:
            print(f"PASS: Component 2 — All 15 SWITCH formulas have correct grade-to-point mappings (0.3 pts)")
            total_score += 0.3
        elif correct_formulas > 0:
            partial_c2 = round(0.3 * correct_formulas / 15, 4)
            print(f"PARTIAL: Component 2 — {correct_formulas}/15 formulas have correct mappings (partial {partial_c2} pts)")
            if incorrect_cells:
                print(f"  Issues: {incorrect_cells[:3]}")
            if partial_c2 > 0:
                total_score += partial_c2
        else:
            print(f"FAIL: Component 2 — No correctly-mapped SWITCH formulas found")
            if incorrect_cells:
                print(f"  Issues: {incorrect_cells[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Compound check — formulas present AND Column A grades / B1 header intact (0.2 points)
    # This is a compound check: task introduced changes (SWITCH in B) AND no side-effect modifications.
    # FAILS on initial because no SWITCH formula in B2 → gate prevents score > 0.
    try:
        expected_grades = ['A', 'B', 'C', 'A', 'F', 'B', 'D', 'C', 'A', 'B', 'F', 'C', 'D', 'A', 'B']
        expected_b1_header = 'Grade Points'

        # Gate: only award points if SWITCH formulas exist (ensures task was performed)
        b2_val = ws.cell(row=2, column=2).value
        has_switch_in_b = (
            b2_val is not None
            and isinstance(b2_val, str)
            and b2_val.upper().startswith('=SWITCH(')
        )

        if not has_switch_in_b:
            print(f"FAIL: Component 3 — No SWITCH formulas found in B2; data integrity not applicable")
        else:
            b1_val = ws.cell(row=1, column=2).value
            header_ok = (b1_val == expected_b1_header)

            wrong_grades = [
                f"A{i}: expected={expected_grade}, got={repr(ws.cell(row=i, column=1).value)}"
                for i, expected_grade in enumerate(expected_grades, start=2)
                if ws.cell(row=i, column=1).value != expected_grade
            ]
            grades_intact = (len(wrong_grades) == 0)

            if header_ok and grades_intact:
                print(f"PASS: Component 3 — Column A grades intact (15 grades) and B1 header correct (0.2 pts)")
                total_score += 0.2
            elif not header_ok:
                print(f"FAIL: Component 3 — B1 header incorrect: expected={repr(expected_b1_header)}, got={repr(b1_val)}")
            else:
                print(f"FAIL: Component 3 — Column A grades modified: {wrong_grades[:3]}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
