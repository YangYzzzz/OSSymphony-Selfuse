"""
Reward Script: Product launch checklist with timeline
Task ID: calc_wf_048
Domain: libreoffice_calc
Scoring:
  - Duration formulas (NETWORKDAYS): 0.25
  - Days Left formulas (MAX/NETWORKDAYS): 0.20
  - Status IF/Overdue formulas: 0.15
  - Priority formulas: 0.10
  - Conditional formatting rules: 0.15
  - Timeline chart: 0.15
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_048'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Launch Plan sheet exists
    if 'Launch Plan' not in wb.sheetnames:
        print("FAIL: 'Launch Plan' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Launch Plan']

    # Precondition: at least 30 data rows
    if ws.max_row < 31:
        print(f"FAIL: Expected at least 31 rows (header + 30 tasks), found {ws.max_row}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Duration column (E) uses NETWORKDAYS formulas (0.25 points)
    try:
        formula_count = 0
        networkdays_count = 0
        for r in range(2, 32):
            val = ws.cell(row=r, column=5).value
            if val and isinstance(val, str) and val.startswith('='):
                formula_count += 1
                if 'NETWORKDAYS' in val.upper():
                    networkdays_count += 1

        if networkdays_count >= 25:
            print(f"PASS: Component 1 — Duration NETWORKDAYS formulas: {networkdays_count}/30 (0.25 pts)")
            total_score += 0.25
        elif networkdays_count >= 15:
            partial = 0.15
            print(f"PARTIAL: Component 1 — Duration NETWORKDAYS formulas: {networkdays_count}/30 ({partial} pts)")
            total_score += partial
        elif formula_count >= 15:
            partial = 0.1
            print(f"PARTIAL: Component 1 — Duration has formulas but not NETWORKDAYS: {formula_count}/30 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Duration formulas: {formula_count}/30, NETWORKDAYS: {networkdays_count}/30")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Days Left column (H) uses MAX/NETWORKDAYS formulas (0.20 points)
    try:
        formula_count = 0
        correct_pattern = 0
        for r in range(2, 32):
            val = ws.cell(row=r, column=8).value
            if val and isinstance(val, str) and val.startswith('='):
                formula_count += 1
                val_upper = val.upper()
                if 'NETWORKDAYS' in val_upper and 'TODAY' in val_upper:
                    correct_pattern += 1

        if correct_pattern >= 25:
            print(f"PASS: Component 2 — Days Left NETWORKDAYS(TODAY) formulas: {correct_pattern}/30 (0.20 pts)")
            total_score += 0.20
        elif correct_pattern >= 15:
            partial = 0.12
            print(f"PARTIAL: Component 2 — Days Left formulas: {correct_pattern}/30 ({partial} pts)")
            total_score += partial
        elif formula_count >= 15:
            partial = 0.08
            print(f"PARTIAL: Component 2 — Days Left has formulas but wrong pattern: {formula_count}/30 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Days Left formulas: {formula_count}/30, correct pattern: {correct_pattern}/30")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Status column (G) with IF/Overdue logic (0.15 points)
    try:
        if_formula_count = 0
        overdue_logic_count = 0
        for r in range(2, 32):
            val = ws.cell(row=r, column=7).value
            if val and isinstance(val, str) and val.startswith('='):
                if_formula_count += 1
                val_upper = val.upper()
                if 'OVERDUE' in val_upper or ('TODAY' in val_upper and 'IF' in val_upper):
                    overdue_logic_count += 1

        # Not all rows need IF formulas - Complete rows may just have the literal value
        if overdue_logic_count >= 10:
            print(f"PASS: Component 3 — Status IF/Overdue formulas: {overdue_logic_count} rows (0.15 pts)")
            total_score += 0.15
        elif if_formula_count >= 5:
            partial = 0.08
            print(f"PARTIAL: Component 3 — Status IF formulas: {if_formula_count}, with Overdue logic: {overdue_logic_count} ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Status IF formulas: {if_formula_count}, Overdue logic: {overdue_logic_count}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Priority column (I) has formulas (0.10 points)
    try:
        formula_count = 0
        if_formula_count = 0
        for r in range(2, 32):
            val = ws.cell(row=r, column=9).value
            if val and isinstance(val, str) and val.startswith('='):
                formula_count += 1
                if 'IF' in val.upper():
                    if_formula_count += 1

        if if_formula_count >= 25:
            print(f"PASS: Component 4 — Priority IF formulas: {if_formula_count}/30 (0.10 pts)")
            total_score += 0.10
        elif formula_count >= 15:
            partial = 0.06
            print(f"PARTIAL: Component 4 — Priority formulas: {formula_count}/30 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Priority formulas: {formula_count}/30")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Conditional formatting rules for status (0.15 points)
    try:
        cf_rules = list(ws.conditional_formatting)
        cf_count = 0
        has_complete_rule = False
        has_overdue_rule = False
        has_in_progress_rule = False
        has_not_started_rule = False

        for cf in cf_rules:
            for rule in cf.rules:
                cf_count += 1
                # Check if rule references status values
                rule_str = str(rule.formula) if hasattr(rule, 'formula') else ''
                rule_str += str(getattr(rule, 'text', '') or '')
                if hasattr(rule, 'formula') and rule.formula:
                    for f in rule.formula:
                        rule_str += str(f)

                rule_str_upper = rule_str.upper()
                if 'COMPLETE' in rule_str_upper and 'NOT' not in rule_str_upper and 'IN' not in rule_str_upper:
                    has_complete_rule = True
                if 'OVERDUE' in rule_str_upper:
                    has_overdue_rule = True
                if 'IN PROGRESS' in rule_str_upper:
                    has_in_progress_rule = True
                if 'NOT STARTED' in rule_str_upper:
                    has_not_started_rule = True

        status_rules_found = sum([has_complete_rule, has_overdue_rule, has_in_progress_rule, has_not_started_rule])

        if status_rules_found >= 4:
            print(f"PASS: Component 5 — All 4 status CF rules found (Complete, In Progress, Not Started, Overdue) (0.15 pts)")
            total_score += 0.15
        elif status_rules_found >= 2:
            partial = 0.08
            print(f"PARTIAL: Component 5 — {status_rules_found}/4 status CF rules found ({partial} pts)")
            total_score += partial
        elif cf_count >= 1:
            partial = 0.05
            print(f"PARTIAL: Component 5 — {cf_count} CF rules found but couldn't match status values ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — No conditional formatting rules found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Timeline chart exists (0.15 points)
    try:
        charts = ws._charts
        if len(charts) >= 1:
            chart = charts[0]
            chart_type = getattr(chart, 'type', None) or getattr(chart, 'tagName', 'unknown')
            chart_title = chart.title if hasattr(chart, 'title') else None
            has_data = len(chart.series) > 0 if hasattr(chart, 'series') else False

            if has_data:
                print(f"PASS: Component 6 — Chart found: type={chart_type}, title={chart_title}, series={len(chart.series)} (0.15 pts)")
                total_score += 0.15
            else:
                partial = 0.08
                print(f"PARTIAL: Component 6 — Chart found but no data series ({partial} pts)")
                total_score += partial
        else:
            print(f"FAIL: Component 6 — No charts found in 'Launch Plan' sheet")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
