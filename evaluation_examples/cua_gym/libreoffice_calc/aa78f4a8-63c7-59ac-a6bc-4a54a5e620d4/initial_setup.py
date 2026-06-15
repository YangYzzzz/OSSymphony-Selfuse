"""
Initial Setup: Product launch checklist with timeline
Task ID: calc_wf_048
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_048'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Launch Plan'

    # Headers
    headers = ['Task', 'Owner', 'Start Date', 'Due Date', 'Duration',
               'Predecessor', 'Status', 'Days Left', 'Priority']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_side = Side(style='thin', color='000000')
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12

    # Base date for the 8-week span
    base_date = date(2026, 4, 6)  # Monday, start of week 1

    # 30 tasks with realistic product launch data
    tasks = [
        # Task, Owner, Start Offset (days), Duration (workdays), Predecessor, Status
        ('Define product vision and goals', 'Sarah Chen', 0, 3, '', 'Complete'),
        ('Conduct market research analysis', 'David Park', 0, 5, '', 'Complete'),
        ('Identify target audience segments', 'Lisa Wong', 3, 4, '1', 'Complete'),
        ('Competitive analysis report', 'David Park', 3, 5, '2', 'Complete'),
        ('Draft product requirements document', 'Sarah Chen', 5, 4, '1,3', 'Complete'),
        ('Create wireframes and mockups', 'James Miller', 7, 6, '5', 'Complete'),
        ('Review and approve designs', 'Sarah Chen', 13, 2, '6', 'Complete'),
        ('Develop product prototype', 'Ryan Torres', 10, 8, '6', 'In Progress'),
        ('Set up testing environment', 'Aisha Patel', 14, 3, '7', 'In Progress'),
        ('Write test cases and scripts', 'Aisha Patel', 14, 4, '7', 'In Progress'),
        ('Internal alpha testing', 'Aisha Patel', 18, 5, '8,9,10', 'Not Started'),
        ('Fix critical bugs from alpha', 'Ryan Torres', 23, 4, '11', 'Not Started'),
        ('Beta testing with focus group', 'Lisa Wong', 27, 5, '12', 'Not Started'),
        ('Collect and analyze beta feedback', 'Lisa Wong', 32, 3, '13', 'Not Started'),
        ('Final product refinements', 'Ryan Torres', 35, 4, '14', 'Not Started'),
        ('Create marketing strategy', 'Marcus Johnson', 10, 5, '3,4', 'In Progress'),
        ('Design brand assets and collateral', 'Emma Davis', 15, 6, '16', 'Not Started'),
        ('Write press release draft', 'Marcus Johnson', 15, 3, '16', 'In Progress'),
        ('Develop social media campaign', 'Emma Davis', 18, 5, '17', 'Not Started'),
        ('Create product landing page', 'James Miller', 20, 5, '17', 'Not Started'),
        ('Set up email marketing sequence', 'Marcus Johnson', 20, 4, '18', 'Not Started'),
        ('Prepare sales training materials', 'Tom Anderson', 25, 4, '16', 'Not Started'),
        ('Conduct sales team training', 'Tom Anderson', 29, 3, '22', 'Not Started'),
        ('Set up customer support workflows', 'Nina Garcia', 25, 5, '9', 'Not Started'),
        ('Create FAQ and help documentation', 'Nina Garcia', 30, 4, '24', 'Not Started'),
        ('Production environment deployment', 'Ryan Torres', 35, 3, '15', 'Not Started'),
        ('Final QA sign-off', 'Aisha Patel', 38, 2, '26', 'Not Started'),
        ('Launch day preparation checklist', 'Sarah Chen', 38, 2, '27', 'Not Started'),
        ('Product launch execution', 'Sarah Chen', 40, 1, '28,23,25', 'Not Started'),
        ('Post-launch monitoring and review', 'Sarah Chen', 41, 4, '29', 'Not Started'),
    ]

    data_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    date_align = Alignment(horizontal='center')

    for i, (task_name, owner, start_offset, duration_days, predecessor, status) in enumerate(tasks, 2):
        start = base_date + timedelta(days=start_offset)
        # Adjust for weekends
        while start.weekday() >= 5:
            start += timedelta(days=1)

        # Calculate due date based on workdays
        due = start
        remaining = duration_days - 1
        while remaining > 0:
            due += timedelta(days=1)
            if due.weekday() < 5:
                remaining -= 1

        row_data = [
            task_name,
            owner,
            start,
            due,
            duration_days,       # Raw number, no formula
            predecessor,
            status,
            '',                  # Days Left - empty, to be filled by formulas
            '',                  # Priority - empty, to be filled by formulas
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = data_border
            if col in (3, 4):  # Date columns
                cell.number_format = 'yyyy-mm-dd'
                cell.alignment = date_align
            elif col in (5, 8):  # Numeric columns
                cell.alignment = Alignment(horizontal='center')
            elif col in (6, 7):  # Center-aligned text columns
                cell.alignment = Alignment(horizontal='center')
            elif col == 9:
                cell.alignment = Alignment(horizontal='center')

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
