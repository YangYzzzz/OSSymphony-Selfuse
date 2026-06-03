"""
Initial Setup: Create HR directory spreadsheet with 100 employee records
Task ID: calc_gg2_009
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg2_009'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    # Headers
    headers = [
        "Employee ID", "First Name", "Last Name",
        "Title", "Location", "Manager", "Department"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic data pools
    first_names = [
        "Sarah", "Marcus", "Elena", "James", "Priya", "David", "Mei",
        "Carlos", "Aisha", "Robert", "Yuki", "Thomas", "Fatima", "Daniel",
        "Olivia", "Raj", "Sophia", "Michael", "Amara", "William",
        "Chen", "Laura", "Ahmed", "Jessica", "Kevin", "Nadia", "Brian",
        "Zara", "Patrick", "Hannah", "Leo", "Valentina", "Nathan",
        "Isabella", "Ryan", "Aaliyah", "Derek", "Camila", "Joshua",
        "Mila", "Andre", "Grace", "Vincent", "Leah", "Gabriel",
        "Naomi", "Oscar", "Samantha", "Felix", "Diana",
    ]

    last_names = [
        "Chen", "Johnson", "Petrov", "Williams", "Sharma", "Kim",
        "Martinez", "Okafor", "Brown", "Tanaka", "Muller", "Davis",
        "Ali", "Anderson", "Nakamura", "Fernandez", "Thompson",
        "Gupta", "Taylor", "Rivera", "Kowalski", "Wilson", "Patel",
        "Garcia", "Lee", "Scott", "Novak", "Hall", "Hassan", "Young",
        "Johansson", "Mitchell", "Ivanova", "Adams", "Park",
        "Lopez", "Turner", "Sato", "Wright", "Rossi",
        "Bell", "Reed", "Price", "Morris", "Ward",
        "Cooper", "Jenkins", "Gray", "Kelly", "Hart",
    ]

    titles = [
        "Software Engineer", "Senior Software Engineer", "Staff Engineer",
        "Engineering Manager", "Product Manager", "Senior Product Manager",
        "Data Analyst", "Senior Data Analyst", "Data Scientist",
        "UX Designer", "Senior UX Designer", "UX Researcher",
        "Marketing Specialist", "Marketing Manager", "Content Strategist",
        "Account Executive", "Sales Manager", "Business Development Rep",
        "HR Coordinator", "HR Manager", "Recruiter",
        "Financial Analyst", "Senior Financial Analyst", "Controller",
        "Operations Manager", "Operations Analyst", "Project Manager",
        "QA Engineer", "Senior QA Engineer", "DevOps Engineer",
    ]

    locations = [
        "San Francisco, CA", "New York, NY", "Austin, TX",
        "Seattle, WA", "Chicago, IL", "Boston, MA",
        "Denver, CO", "Los Angeles, CA", "Portland, OR",
        "Atlanta, GA",
    ]

    departments = [
        "Engineering", "Product", "Data Science",
        "Design", "Marketing", "Sales",
        "Human Resources", "Finance", "Operations",
    ]

    managers = [
        "Jennifer Walsh", "Mark Sullivan", "Lisa Huang",
        "Robert Pierce", "Amanda Foster", "Steven Clark",
        "Diane Reyes", "Gregory Barnes", "Monica Chavez",
        "Philip Lawson",
    ]

    random.seed(42)  # Reproducible data

    for i in range(100):
        row = i + 2
        emp_id = f"EMP-{1001 + i}"
        first = random.choice(first_names)
        last = random.choice(last_names)
        title = random.choice(titles)
        location = random.choice(locations)
        manager = random.choice(managers)
        dept = random.choice(departments)

        ws.cell(row=row, column=1, value=emp_id)
        ws.cell(row=row, column=2, value=first)
        ws.cell(row=row, column=3, value=last)
        ws.cell(row=row, column=4, value=title)
        ws.cell(row=row, column=5, value=location)
        ws.cell(row=row, column=6, value=manager)
        ws.cell(row=row, column=7, value=dept)

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18

    # NO conditional formatting in initial state
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
