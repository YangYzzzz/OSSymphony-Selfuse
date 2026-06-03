"""
Reward Script: Apply thick box border, dark blue fill, and white font to header row B2:F2
Task ID: calc_gg3_033
Domain: libreoffice_calc
Scoring:
  Component 1 — Dark blue background (#003366) on B2:F2  (0.35 pts)
  Component 2 — White font color (#FFFFFF) on B2:F2      (0.35 pts)
  Component 3 — Thick box border around perimeter of B2:F2 (0.30 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_033'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Table' sheet must exist
    if 'Table' not in wb.sheetnames:
        print("FAIL: 'Table' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Table']

    # Header cells to check: B2 through F2 (columns 2-6, row 2)
    header_cells = [ws.cell(row=2, column=c) for c in range(2, 7)]
    header_coords = ['B2', 'C2', 'D2', 'E2', 'F2']

    # Component 1: Dark blue background fill (#003366) on all 5 header cells (0.35 pts)
    # Expected: PatternFill solid with fgColor = FF003366 (ARGB)
    try:
        fill_pass_count = 0
        for cell, coord in zip(header_cells, header_coords):
            try:
                fg = cell.fill.fgColor.rgb if cell.fill.fgColor else None
                fill_type = cell.fill.fill_type
            except Exception:
                fg = None
                fill_type = None

            # Accept FF003366 as the expected dark blue ARGB
            if fill_type == 'solid' and fg and fg.upper() == 'FF003366':
                fill_pass_count += 1
            else:
                print(f"  DETAIL: {coord} fill: type={fill_type}, fgColor={fg}")

        if fill_pass_count == 5:
            print(f"PASS: Component 1 — All 5 header cells have dark blue background (0.35 pts)")
            total_score += 0.35
        elif fill_pass_count > 0:
            partial = round(0.35 * fill_pass_count / 5, 2)
            print(f"PARTIAL: Component 1 — {fill_pass_count}/5 cells have dark blue fill ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No header cells have dark blue background fill")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: White font color (#FFFFFF) on all 5 header cells (0.35 pts)
    # Expected: font.color.rgb contains FFFFFF (may be 00FFFFFF or FFFFFFFF)
    try:
        font_pass_count = 0
        for cell, coord in zip(header_cells, header_coords):
            try:
                font_rgb = cell.font.color.rgb if cell.font.color else None
            except Exception:
                font_rgb = None

            # Accept any ARGB ending in FFFFFF (white)
            if font_rgb and isinstance(font_rgb, str) and font_rgb.upper().endswith('FFFFFF'):
                font_pass_count += 1
            else:
                print(f"  DETAIL: {coord} font color: {font_rgb}")

        if font_pass_count == 5:
            print(f"PASS: Component 2 — All 5 header cells have white font color (0.35 pts)")
            total_score += 0.35
        elif font_pass_count > 0:
            partial = round(0.35 * font_pass_count / 5, 2)
            print(f"PARTIAL: Component 2 — {font_pass_count}/5 cells have white font ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No header cells have white font color")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Thick box border around the perimeter of B2:F2 (0.30 pts)
    # A "thick box border" means:
    #   - All cells: top=thick, bottom=thick
    #   - B2 (leftmost): left=thick
    #   - F2 (rightmost): right=thick
    # Interior vertical borders (right of B2..E2, left of C2..F2) are not required by "box border"
    try:
        border_checks_pass = 0
        total_border_checks = 0

        for i, (cell, coord) in enumerate(zip(header_cells, header_coords)):
            b = cell.border

            # Top must be thick
            total_border_checks += 1
            if b.top.style == 'thick':
                border_checks_pass += 1
            else:
                print(f"  DETAIL: {coord} top border: {b.top.style} (expected thick)")

            # Bottom must be thick
            total_border_checks += 1
            if b.bottom.style == 'thick':
                border_checks_pass += 1
            else:
                print(f"  DETAIL: {coord} bottom border: {b.bottom.style} (expected thick)")

            # Left border: only required on B2 (first cell)
            if i == 0:
                total_border_checks += 1
                if b.left.style == 'thick':
                    border_checks_pass += 1
                else:
                    print(f"  DETAIL: {coord} left border: {b.left.style} (expected thick)")

            # Right border: only required on F2 (last cell)
            if i == 4:
                total_border_checks += 1
                if b.right.style == 'thick':
                    border_checks_pass += 1
                else:
                    print(f"  DETAIL: {coord} right border: {b.right.style} (expected thick)")

        if border_checks_pass == total_border_checks:
            print(f"PASS: Component 3 — Thick box border on B2:F2 perimeter ({border_checks_pass}/{total_border_checks} checks) (0.30 pts)")
            total_score += 0.30
        elif border_checks_pass > 0:
            partial = round(0.30 * border_checks_pass / total_border_checks, 2)
            print(f"PARTIAL: Component 3 — {border_checks_pass}/{total_border_checks} border checks pass ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No thick border found on B2:F2 perimeter")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
