"""
Initial Setup: Apply thick box border, dark blue fill, and white font to header row B2:F2
Task ID: calc_gg3_033
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_033'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Table'

    # Headers in row 2, columns B-F (no formatting — plain text)
    headers = ['Name', 'Department', 'Start Date', 'Salary', 'Status']
    for col_idx, h in enumerate(headers, 2):  # B=2, C=3, D=4, E=5, F=6
        ws.cell(row=2, column=col_idx, value=h)

    # 50 rows of realistic employee data (rows 3-52)
    employees = [
        ['Sarah Chen', 'Engineering', '2023-01-15', 92000, 'Active'],
        ['Marcus Johnson', 'Marketing', '2022-06-01', 78500, 'Active'],
        ['Emily Rodriguez', 'Finance', '2021-09-20', 85000, 'Active'],
        ['James Kim', 'Engineering', '2024-02-10', 88000, 'Active'],
        ['Olivia Brown', 'Human Resources', '2020-11-05', 72000, 'On Leave'],
        ['Daniel Martinez', 'Sales', '2023-04-18', 67500, 'Active'],
        ['Sophia Williams', 'Engineering', '2022-08-22', 95000, 'Active'],
        ['Liam Davis', 'Marketing', '2021-03-30', 71000, 'Active'],
        ['Ava Thompson', 'Finance', '2023-07-12', 82000, 'Active'],
        ['Noah Garcia', 'Operations', '2020-05-14', 69000, 'Terminated'],
        ['Isabella Wilson', 'Engineering', '2024-01-08', 91000, 'Active'],
        ['Mason Anderson', 'Sales', '2022-10-25', 73500, 'Active'],
        ['Mia Taylor', 'Human Resources', '2021-06-17', 68000, 'Active'],
        ['Ethan Thomas', 'Finance', '2023-09-03', 87000, 'Active'],
        ['Amelia Jackson', 'Marketing', '2020-12-19', 76000, 'On Leave'],
        ['Alexander White', 'Engineering', '2022-04-07', 93500, 'Active'],
        ['Charlotte Harris', 'Operations', '2021-08-11', 65000, 'Active'],
        ['Benjamin Clark', 'Sales', '2023-03-28', 70000, 'Active'],
        ['Harper Lewis', 'Finance', '2024-05-15', 84000, 'Active'],
        ['William Robinson', 'Engineering', '2022-01-20', 96000, 'Active'],
        ['Evelyn Walker', 'Human Resources', '2020-07-09', 71500, 'Active'],
        ['Henry Young', 'Marketing', '2023-11-14', 74000, 'Active'],
        ['Abigail Allen', 'Sales', '2021-02-26', 68500, 'Active'],
        ['Sebastian King', 'Operations', '2022-12-08', 67000, 'Terminated'],
        ['Ella Wright', 'Engineering', '2024-03-19', 89000, 'Active'],
        ['Jack Scott', 'Finance', '2021-05-22', 83000, 'Active'],
        ['Scarlett Green', 'Marketing', '2023-06-30', 72500, 'Active'],
        ['Owen Baker', 'Sales', '2020-09-15', 69500, 'Active'],
        ['Grace Adams', 'Human Resources', '2022-07-04', 70000, 'On Leave'],
        ['Lucas Nelson', 'Engineering', '2023-10-21', 94000, 'Active'],
        ['Chloe Hill', 'Operations', '2021-04-13', 66000, 'Active'],
        ['Aiden Campbell', 'Finance', '2024-06-01', 86000, 'Active'],
        ['Zoey Mitchell', 'Marketing', '2022-03-17', 75000, 'Active'],
        ['Logan Roberts', 'Sales', '2020-08-29', 71000, 'Active'],
        ['Lily Carter', 'Engineering', '2023-02-05', 90000, 'Active'],
        ['Matthew Phillips', 'Human Resources', '2021-11-18', 69000, 'Active'],
        ['Aria Evans', 'Finance', '2022-09-24', 81000, 'Active'],
        ['Jackson Turner', 'Operations', '2024-04-08', 64500, 'Active'],
        ['Riley Torres', 'Marketing', '2023-08-16', 73000, 'Active'],
        ['Gabriel Parker', 'Sales', '2021-01-27', 67500, 'Terminated'],
        ['Layla Collins', 'Engineering', '2022-05-11', 92500, 'Active'],
        ['David Edwards', 'Finance', '2020-10-03', 85500, 'Active'],
        ['Penelope Stewart', 'Human Resources', '2023-12-22', 70500, 'Active'],
        ['Carter Sanchez', 'Operations', '2022-02-14', 66500, 'Active'],
        ['Nora Morris', 'Marketing', '2021-07-08', 74500, 'On Leave'],
        ['Wyatt Rogers', 'Engineering', '2024-07-10', 88500, 'Active'],
        ['Hannah Reed', 'Sales', '2023-05-25', 69000, 'Active'],
        ['Dylan Cook', 'Finance', '2022-11-30', 83500, 'Active'],
        ['Addison Morgan', 'Operations', '2021-10-06', 65500, 'Active'],
        ['Luke Bell', 'Engineering', '2020-04-22', 97000, 'Active'],
    ]

    for r, emp in enumerate(employees, 3):
        ws.cell(row=r, column=2, value=emp[0])  # Name
        ws.cell(row=r, column=3, value=emp[1])  # Department
        ws.cell(row=r, column=4, value=emp[2])  # Start Date
        ws.cell(row=r, column=5, value=emp[3])  # Salary
        ws.cell(row=r, column=6, value=emp[4])  # Status

    # Set reasonable column widths
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
