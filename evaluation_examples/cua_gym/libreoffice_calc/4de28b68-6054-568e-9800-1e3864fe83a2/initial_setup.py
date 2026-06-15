"""
Initial Setup: Add polynomial trendline to scatter chart
Task ID: calc_chart_trendline_polynomial_060
Domain: libreoffice_calc

Creates a spreadsheet with:
- Sheet 'MemoryStudy' containing age and memory score data
- A scatter chart titled 'Age vs Memory Score' (NO trendline initially)
"""

import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_trendline_polynomial_060'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: MemoryStudy ---
    ws = wb.active
    ws.title = 'MemoryStudy'

    # Headers
    ws['A1'] = 'Age'
    ws['B1'] = 'Memory Score'

    # Data rows as specified in context
    data = [
        (20, 88),
        (25, 91),
        (30, 89),
        (35, 85),
        (40, 80),
        (45, 74),
        (50, 69),
        (55, 63),
        (60, 58),
        (65, 54),
    ]

    for r, (age, score) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=age)
        ws.cell(row=r, column=2, value=score)

    # Style the header row
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    for col in ['A', 'B']:
        ws[f'{col}1'].font = header_font
        ws[f'{col}1'].fill = header_fill
        ws[f'{col}1'].alignment = Alignment(horizontal='center')

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 16

    # --- Create Scatter Chart (NO trendline) ---
    scatter = ScatterChart()
    scatter.title = 'Age vs Memory Score'
    scatter.style = 10
    scatter.x_axis.title = 'Age'
    scatter.y_axis.title = 'Memory Score'
    scatter.x_axis.crosses = 'autoZero'
    scatter.x_axis.numFmt = '0'
    scatter.y_axis.numFmt = '0'

    # X values: Age column (A2:A11)
    x_vals = Reference(ws, min_col=1, min_row=2, max_row=11)
    # Y values: Memory Score column (B2:B11)
    y_vals = Reference(ws, min_col=2, min_row=2, max_row=11)

    series = Series(y_vals, x_vals, title='Memory Score')
    scatter.series.append(series)

    # Place chart on the sheet
    ws.add_chart(scatter, 'D2')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('  - Sheet: MemoryStudy')
    print('  - Data: 10 rows of Age vs Memory Score data')
    print('  - Chart: Scatter chart "Age vs Memory Score" (no trendline)')


create_initial()
