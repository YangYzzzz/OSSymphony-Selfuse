"""
Initial Setup: Create quality control spreadsheet with metrics and thresholds
Task ID: calc_gg2_046
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg2_046'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quality Control"

    # --- Headers ---
    headers = [
        "Product", "Metric1", "Metric2", "Metric3", "Metric4", "Metric5",
        "Threshold1", "Threshold2", "Threshold3", "Threshold4", "Threshold5"
    ]
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Product names (25 realistic products) ---
    products = [
        "AeroBlend Pro X1", "TurboMax 500", "PrecisionCut Elite",
        "HydraFlow Controller", "NanoCoat Shield V2", "FlexiGrip Handle",
        "PowerSync Module", "ClearView Lens Kit", "ThermoLock Seal",
        "VoltEdge Connector", "SilkTouch Panel", "IronClad Bracket",
        "SpeedLink Adapter", "AquaPure Filter", "MegaDrive Motor",
        "SmartSense Probe", "TitanFrame Chassis", "EcoVolt Regulator",
        "OptiFlow Valve", "DigiTrace Sensor", "RapidBond Adhesive",
        "StealthCoat Film", "ZenithGear Pulley", "CoreBalance Rotor",
        "PeakPerform Actuator"
    ]

    # Seed for reproducibility
    random.seed(42)

    # Generate metric values and thresholds
    # Metrics represent quality scores (e.g., tensile strength, purity %, defect rate, etc.)
    for r, product in enumerate(products, 2):
        ws.cell(row=r, column=1, value=product).border = thin_border

        for metric_col in range(2, 7):  # B through F (Metric1-Metric5)
            # Generate realistic metric values between 60 and 100
            value = round(random.uniform(60, 100), 1)
            cell = ws.cell(row=r, column=metric_col, value=value)
            cell.border = thin_border
            cell.number_format = '0.0'
            cell.alignment = Alignment(horizontal="center")

        for thresh_col in range(7, 12):  # G through K (Threshold1-Threshold5)
            # Thresholds between 70 and 90, so some metrics will fail
            threshold = round(random.uniform(70, 90), 1)
            cell = ws.cell(row=r, column=thresh_col, value=threshold)
            cell.border = thin_border
            cell.number_format = '0.0'
            cell.alignment = Alignment(horizontal="center")

    # --- Column widths ---
    ws.column_dimensions['A'].width = 22
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        ws.column_dimensions[col_letter].width = 14

    # --- Freeze header row ---
    ws.freeze_panes = "A2"

    # NO conditional formatting in initial state (task requires agent to add it)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
