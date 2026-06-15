"""
Reward Script: Conditional formatting for quality control metrics
Task ID: calc_gg2_046
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.2): At least 5 conditional formatting rules exist
  - Components 2-6 (0.16 each): Each of the 5 column rules has correct formula,
    range, and red fill color
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg2_046'

# Expected rules: column letter -> (formula, range_prefix)
EXPECTED_RULES = {
    'B': ('$B2<$G2', 'B2:B26'),
    'C': ('$C2<$H2', 'C2:C26'),
    'D': ('$D2<$I2', 'D2:D26'),
    'E': ('$E2<$J2', 'E2:E26'),
    'F': ('$F2<$K2', 'F2:F26'),
}

RED_ARGB = 'FFFF0000'


def normalize_formula(f):
    """Normalize formula for comparison (uppercase, no spaces)."""
    return f.upper().replace(' ', '')


def check_range_covers_column(cf_range_str, col_letter):
    """
    Check if the conditional formatting range covers the expected column rows 2-26.
    Accepts exact match like 'B2:B26' or broader ranges like 'B2:F26' that include the column.
    """
    from openpyxl.utils import column_index_from_string, get_column_letter
    range_str = str(cf_range_str).strip()

    # Parse range like "B2:F26" or "B2:B26"
    try:
        parts = range_str.split(':')
        if len(parts) != 2:
            return False

        # Extract start col/row and end col/row
        start = parts[0]
        end = parts[1]

        # Parse column letters and row numbers
        start_col = ''
        start_row = ''
        for ch in start:
            if ch.isalpha():
                start_col += ch
            else:
                start_row += ch

        end_col = ''
        end_row = ''
        for ch in end:
            if ch.isalpha():
                end_col += ch
            else:
                end_row += ch

        start_col_idx = column_index_from_string(start_col.upper())
        end_col_idx = column_index_from_string(end_col.upper())
        target_col_idx = column_index_from_string(col_letter.upper())

        start_row_num = int(start_row)
        end_row_num = int(end_row)

        # The range must include the target column and cover rows 2-26
        if target_col_idx < start_col_idx or target_col_idx > end_col_idx:
            return False
        if start_row_num > 2 or end_row_num < 26:
            return False

        return True
    except Exception:
        return False


def verify_task(file_path):
    """
    Verify conditional formatting rules for quality control metrics.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check 'Quality Control' sheet exists
    if 'Quality Control' not in wb.sheetnames:
        print("FAIL: 'Quality Control' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Quality Control']

    # Collect all conditional formatting rules with their ranges and details
    cf_rules_found = []
    for cf in ws.conditional_formatting:
        # str(cf) returns '<ConditionalFormatting B2:B26>' — extract the cell range
        raw_range = str(cf)
        # Extract range from angle brackets if present
        if '>' in raw_range:
            range_str = raw_range.split()[-1].rstrip('>')
        else:
            range_str = raw_range
        for rule in cf.rules:
            formula_list = rule.formula if rule.formula else []
            formula_str = formula_list[0] if formula_list else ''
            # Check fill color
            fill_color = None
            if rule.dxf and rule.dxf.fill and rule.dxf.fill.fgColor:
                try:
                    fill_color = rule.dxf.fill.fgColor.rgb
                except Exception:
                    pass
            cf_rules_found.append({
                'range': range_str,
                'type': rule.type,
                'formula': formula_str,
                'fill_color': fill_color,
            })

    print(f"Found {len(cf_rules_found)} conditional formatting rules total")
    for i, r in enumerate(cf_rules_found):
        print(f"  Rule {i+1}: range={r['range']}, type={r['type']}, "
              f"formula={r['formula']}, fill={r['fill_color']}")

    # Component 1: At least 5 conditional formatting rules exist (0.2 points)
    try:
        if len(cf_rules_found) >= 5:
            print(f"PASS: Component 1 -- {len(cf_rules_found)} CF rules found >= 5 (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 1 -- only {len(cf_rules_found)} CF rules, need >= 5")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Components 2-6: Verify each column's rule (0.16 points each)
    component_num = 2
    for col_letter, (expected_formula, expected_range) in EXPECTED_RULES.items():
        try:
            matched = False
            for rule_info in cf_rules_found:
                # Check formula matches
                if normalize_formula(rule_info['formula']) != normalize_formula(expected_formula):
                    continue
                # Check range covers the expected column
                if not check_range_covers_column(rule_info['range'], col_letter):
                    continue
                # Check rule type is expression/formula
                if rule_info['type'] not in ('expression', 'formula'):
                    continue
                # Check fill color is red
                if rule_info['fill_color'] and 'FF0000' in rule_info['fill_color'].upper():
                    matched = True
                    break

            if matched:
                print(f"PASS: Component {component_num} -- Column {col_letter} rule "
                      f"with formula {expected_formula} and red fill (0.16 pts)")
                total_score += 0.16
            else:
                print(f"FAIL: Component {component_num} -- Column {col_letter} rule "
                      f"not found or incorrect (expected formula={expected_formula}, "
                      f"range covering {expected_range}, red fill)")
        except Exception as e:
            print(f"ERROR: Component {component_num} -- {e}")
        component_num += 1

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved LibreOffice edits before verification
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Main entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
