"""
Reward Script: Extract EU Horizon Europe Marie Curie fellowship statistics from PDFs
Task ID: osworld_multi_apps_pdf_stats_table_008
Domain: libreoffice_calc
Scoring:
  Component 1: File exists at /home/user/Desktop/Marie_Curie_rates.xlsx with correct 4-column headers (0.25 pts)
  Component 2: All 5 data rows (2020-2024) with correct applications and selected counts (0.40 pts)
  Component 3: All 5 selection rates computed correctly to 2 decimal places (0.35 pts)
Total: 1.0
"""

import os

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_pdf_stats_table_008'
FILE_PATH = '/home/user/Desktop/Marie_Curie_rates.xlsx'

# Ground truth values extracted from golden_env
# Year -> (Applications, Selected, Selection_Rate_2dp)
EXPECTED_DATA = {
    2020: (11185, 1610, 14.39),
    2021: (8880,  1270, 14.30),
    2022: (9614,  1437, 14.95),
    2023: (10342, 1523, 14.73),
    2024: (10891, 1598, 14.67),
}

EXPECTED_HEADERS = ['Year', 'Applications', 'Selected', 'Selection Rate (%)']


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Gate: file must exist — not scored, just a prerequisite
    if not os.path.exists(file_path):
        print(f"FAIL (gate): File not found at {file_path}")
        print("REWARD: 0.0")
        return 0.0

    # Import openpyxl only after confirming the file exists
    import openpyxl

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Find the active/first worksheet
    try:
        ws = wb.active
    except Exception as e:
        print(f"CRITICAL: Cannot access worksheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # -----------------------------------------------------------------------
    # Component 1: Correct 4-column headers in row 1 (0.25 points)
    # This FAILS on initial_env (file does not exist there) and PASSES on golden_env.
    # -----------------------------------------------------------------------
    try:
        headers = []
        for col_idx in range(1, 5):
            cell_val = ws.cell(row=1, column=col_idx).value
            headers.append(str(cell_val).strip() if cell_val is not None else '')

        # Normalize case-insensitive comparison
        headers_normalized = [h.lower() for h in headers]
        expected_normalized = [h.lower() for h in EXPECTED_HEADERS]

        if headers_normalized == expected_normalized:
            print(f"PASS: Component 1 — Correct 4-column headers found: {headers} (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — Expected headers {EXPECTED_HEADERS}, found {headers}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: All 5 data rows (2020-2024) with correct applications and
    # selected counts (0.40 points)
    # This FAILS on initial_env and PASSES on golden_env.
    # -----------------------------------------------------------------------
    try:
        # Read all data rows (rows 2 onward)
        found_rows = {}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if row[0] is not None:
                try:
                    year = int(row[0])
                    apps = int(row[1]) if row[1] is not None else None
                    sel  = int(row[2]) if row[2] is not None else None
                    found_rows[year] = (apps, sel)
                except (ValueError, TypeError):
                    continue

        years_correct = 0
        years_expected = sorted(EXPECTED_DATA.keys())

        for year in years_expected:
            exp_apps, exp_sel, _ = EXPECTED_DATA[year]
            if year in found_rows:
                got_apps, got_sel = found_rows[year]
                if got_apps == exp_apps and got_sel == exp_sel:
                    years_correct += 1
                    print(f"  PASS: Year {year}: apps={got_apps}, selected={got_sel}")
                else:
                    print(f"  FAIL: Year {year}: expected apps={exp_apps}, selected={exp_sel}; "
                          f"found apps={got_apps}, selected={got_sel}")
            else:
                print(f"  FAIL: Year {year} not found in file (found years: {sorted(found_rows.keys())})")

        if years_correct == 5:
            print(f"PASS: Component 2 — All 5 years have correct Applications and Selected counts (0.40 pts)")
            total_score += 0.40
        elif years_correct >= 3:
            partial = round(0.40 * years_correct / 5, 2)
            print(f"PARTIAL: Component 2 — {years_correct}/5 years correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {years_correct}/5 years have correct counts (0 pts)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: All 5 selection rates computed correctly to 2 decimal places
    # (0.35 points)
    # Selection Rate = round(Selected / Applications * 100, 2)
    # This FAILS on initial_env and PASSES on golden_env.
    # -----------------------------------------------------------------------
    try:
        # Re-read data rows to get rate values
        found_rates = {}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if row[0] is not None and row[3] is not None:
                try:
                    year = int(row[0])
                    rate = float(row[3])
                    found_rates[year] = rate
                except (ValueError, TypeError):
                    continue

        rates_correct = 0
        for year in sorted(EXPECTED_DATA.keys()):
            exp_apps, exp_sel, exp_rate = EXPECTED_DATA[year]
            computed_rate = round(exp_sel / exp_apps * 100, 2)

            if year in found_rates:
                got_rate = found_rates[year]
                # Allow tolerance of 0.005 (rounding at 2dp)
                if abs(got_rate - computed_rate) <= 0.005:
                    rates_correct += 1
                    print(f"  PASS: Year {year}: selection rate={got_rate:.2f}% (expected {computed_rate:.2f}%)")
                else:
                    print(f"  FAIL: Year {year}: selection rate={got_rate}, expected {computed_rate:.2f}%")
            else:
                print(f"  FAIL: Year {year}: no selection rate found in file")

        if rates_correct == 5:
            print(f"PASS: Component 3 — All 5 selection rates correct to 2 decimal places (0.35 pts)")
            total_score += 0.35
        elif rates_correct >= 3:
            partial = round(0.35 * rates_correct / 5, 2)
            print(f"PARTIAL: Component 3 — {rates_correct}/5 rates correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {rates_correct}/5 selection rates are correct (0 pts)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint: verify against canonical artifact path
if not os.path.exists(FILE_PATH):
    print(f"File not found: {FILE_PATH}")
    print("REWARD: 0.0")
else:
    verify_task(FILE_PATH)
