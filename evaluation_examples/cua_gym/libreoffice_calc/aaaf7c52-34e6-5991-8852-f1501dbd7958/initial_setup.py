"""
Initial Setup: Monthly Executive Sales Report
Task ID: calc_sales_report_executive_043
Domain: libreoffice_calc

Creates the pre-task state for the executive sales report spreadsheet.
The file has raw data with no formatting, no merged cells, no formulas in B5,
no charts, and no print settings applied.
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_report_executive_043'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # ---- ExecReport Sheet ----
    ws = wb.active
    ws.title = 'ExecReport'

    # A1: Report title (unformatted, not merged)
    ws['A1'] = 'Monthly Sales Report'

    # Row 2: blank separator
    # A3:B3 - Total Revenue row
    ws['A3'] = 'Total Revenue'
    ws['B3'] = 4250000

    # A4:B4 - Monthly Target row
    ws['A4'] = 'Monthly Target'
    ws['B4'] = 4000000

    # A5:B5 - vs Target row (B5 intentionally empty - task requires formula)
    ws['A5'] = 'vs Target'
    ws['B5'] = None  # Empty - agent must add =(B3-B4)/B4

    # Row 6: blank separator

    # A7:D12 - Top 5 Deals Table (headers in row 7, data in rows 8-12)
    ws['A7'] = 'Deal Name'
    ws['B7'] = 'Account'
    ws['C7'] = 'Deal Value'
    ws['D7'] = 'Close Date'

    top_deals = [
        ('Project Phoenix Expansion',    'Nextech Solutions',     875000,  '2025-02-28'),
        ('Enterprise Suite Renewal',     'GlobalCorp Inc.',       612500,  '2025-02-24'),
        ('Cloud Migration Package',      'Meridian Healthcare',   540000,  '2025-02-21'),
        ('Analytics Platform License',   'Summit Financial',      425000,  '2025-02-19'),
        ('Security Compliance Bundle',   'Pacific Retail Group',  387500,  '2025-02-15'),
    ]
    for r, (deal, account, value, date) in enumerate(top_deals, 8):
        ws.cell(row=r, column=1, value=deal)
        ws.cell(row=r, column=2, value=account)
        ws.cell(row=r, column=3, value=value)
        ws.cell(row=r, column=4, value=date)

    # Row 13: blank separator

    # A14:D18 - Regional Breakdown Table (headers in row 14, data in rows 15-18)
    ws['A14'] = 'Region'
    ws['B14'] = 'Revenue'
    ws['C14'] = 'Target'
    ws['D14'] = 'Attainment'

    regional_data = [
        ('North America',  1987500, 1800000, 0.1042),
        ('EMEA',           1125000, 1100000, 0.0227),
        ('Asia Pacific',    843750,  800000, 0.0547),
        ('Latin America',   293750,  300000, -0.0208),
    ]
    for r, (region, revenue, target, attain) in enumerate(regional_data, 15):
        ws.cell(row=r, column=1, value=region)
        ws.cell(row=r, column=2, value=revenue)
        ws.cell(row=r, column=3, value=target)
        ws.cell(row=r, column=4, value=attain)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets:', wb.sheetnames)
    print('No formatting, no merged cells, no formulas, no charts applied.')


create_initial()
