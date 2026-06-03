"""
Reward Script: Monthly Executive Sales Report Formatting
Task ID: calc_sales_report_executive_043
Domain: libreoffice_calc
Scoring:
  Component 1: A1:D1 merged header with dark blue fill, white bold text, size 16, centered (0.25)
  Component 2: B5 contains =(B3-B4)/B4 formula with percentage number format (0.20)
  Component 3: Currency format on B3, B4, and monetary deal values (0.15)
  Component 4: Dark blue header rows (3, 7, 14) with bold, dark blue fill, white text (0.20)
  Component 5: Bar chart embedded with regional data title (0.10)
  Component 6: Print header with report title/date and footer with page number (0.10)
  Total: 1.0
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_report_executive_043'

DARK_BLUE_FILL = 'FF1F3864'   # Expected dark blue background (ARGB)
WHITE_FONT = 'FFFFFFFF'        # Expected white text (ARGB)


def get_font_color_rgb(cell):
    """Safely extract font color rgb string from a cell, returning None if not available."""
    try:
        fc = cell.font.color
        if fc and hasattr(fc, 'rgb') and isinstance(fc.rgb, str):
            return fc.rgb
    except Exception:
        pass
    return None


def get_fill_color_rgb(cell):
    """Safely extract fill fgColor rgb string from a cell."""
    try:
        return cell.fill.fgColor.rgb
    except Exception:
        return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: file must load
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: ExecReport sheet must exist
    if 'ExecReport' not in wb.sheetnames:
        print("CRITICAL: Sheet 'ExecReport' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['ExecReport']

    # ------------------------------------------------------------------
    # Component 1: Merged header A1:D1 with professional formatting (0.25 pts)
    # - A1:D1 must be merged into one cell
    # - Font: bold=True, size>=14, white color (FFFFFFFF)
    # - Fill: dark blue (FF1F3864)
    # - Alignment: horizontal center
    # Points are accumulated sub-component by sub-component
    # ------------------------------------------------------------------
    try:
        comp1_score = 0.0
        comp1_details = []

        # Sub-check 1a: A1:D1 is merged (0.10 pts)
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        if 'A1:D1' in merged_ranges:
            comp1_score += 0.10
            comp1_details.append("A1:D1 merged")
        else:
            comp1_details.append(f"FAIL: A1:D1 not merged (ranges={merged_ranges})")

        cell_a1 = ws['A1']

        # Sub-check 1b: Bold font (0.05 pts)
        if cell_a1.font.bold:
            comp1_score += 0.05
            comp1_details.append("A1 bold")
        else:
            comp1_details.append("FAIL: A1 not bold")

        # Sub-check 1c: Font size >= 14 (0.04 pts)
        if cell_a1.font.size and cell_a1.font.size >= 14:
            comp1_score += 0.04
            comp1_details.append(f"A1 font size={cell_a1.font.size}")
        else:
            comp1_details.append(f"FAIL: A1 font size={cell_a1.font.size} (expected >=14)")

        # Sub-check 1d: Dark blue fill (0.04 pts)
        fill_color = get_fill_color_rgb(cell_a1)
        if fill_color and fill_color.upper() == DARK_BLUE_FILL.upper():
            comp1_score += 0.04
            comp1_details.append(f"A1 fill dark blue ({fill_color})")
        else:
            comp1_details.append(f"FAIL: A1 fill={fill_color} (expected {DARK_BLUE_FILL})")

        # Sub-check 1e: White font color (0.02 pts)
        font_color = get_font_color_rgb(cell_a1)
        if font_color and font_color.upper() in (WHITE_FONT.upper(), 'FFFFFF'):
            comp1_score += 0.02
            comp1_details.append(f"A1 font color white ({font_color})")
        else:
            comp1_details.append(f"FAIL: A1 font color={font_color} (expected white {WHITE_FONT})")

        if comp1_score >= 0.20:
            print(f"PASS: Component 1 — Merged header formatting ({comp1_score:.2f} pts) — {'; '.join(comp1_details)}")
        else:
            print(f"PARTIAL: Component 1 — Merged header ({comp1_score:.2f}/0.25 pts) — {'; '.join(comp1_details)}")

        # Aggregate sub-component score into total
        if comp1_score > 0:
            total_score += comp1_score

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ------------------------------------------------------------------
    # Component 2: B5 contains =(B3-B4)/B4 formula with percentage format (0.20 pts)
    # ------------------------------------------------------------------
    try:
        comp2_score = 0.0
        comp2_details = []

        cell_b5 = ws['B5']
        b5_value = cell_b5.value

        # Sub-check 2a: formula is present (0.12 pts)
        if b5_value and isinstance(b5_value, str):
            normalized = b5_value.upper().replace(' ', '')
            if '=(B3-B4)/B4' in normalized or normalized == '=(B3-B4)/B4':
                comp2_score += 0.12
                comp2_details.append(f"B5 formula correct ({b5_value})")
            else:
                comp2_details.append(f"FAIL: B5 formula={repr(b5_value)} (expected =(B3-B4)/B4)")
        else:
            comp2_details.append(f"FAIL: B5 value={repr(b5_value)} (expected formula string)")

        # Sub-check 2b: percentage number format on B5 (0.08 pts)
        b5_fmt = cell_b5.number_format
        if b5_fmt and '%' in b5_fmt:
            comp2_score += 0.08
            comp2_details.append(f"B5 number_format={b5_fmt} (percentage)")
        else:
            comp2_details.append(f"FAIL: B5 number_format={b5_fmt} (expected % format)")

        if comp2_score >= 0.16:
            print(f"PASS: Component 2 — MoM growth formula ({comp2_score:.2f} pts) — {'; '.join(comp2_details)}")
        else:
            print(f"PARTIAL: Component 2 — MoM growth formula ({comp2_score:.2f}/0.20 pts) — {'; '.join(comp2_details)}")

        if comp2_score > 0:
            total_score += comp2_score

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ------------------------------------------------------------------
    # Component 3: Currency format on B3, B4 and deal value cells (0.15 pts)
    # Initial file has 'General' number format; golden applies $#,##0.00
    # ------------------------------------------------------------------
    try:
        comp3_details = []
        currency_cells = {'B3': ws['B3'], 'B4': ws['B4'], 'C8': ws['C8']}
        currency_passed = 0

        for coord, cell in currency_cells.items():
            fmt = cell.number_format
            if fmt and '$' in fmt:
                currency_passed += 1
                comp3_details.append(f"{coord} fmt={fmt}")
            else:
                comp3_details.append(f"FAIL {coord} fmt={fmt}")

        # Award 0.05 per currency cell correctly formatted
        comp3_score = currency_passed * 0.05
        if comp3_score >= 0.14:
            print(f"PASS: Component 3 — Currency formatting ({comp3_score:.2f} pts) — {'; '.join(comp3_details)}")
        else:
            print(f"PARTIAL: Component 3 — Currency formatting ({comp3_score:.2f}/0.15 pts) — {'; '.join(comp3_details)}")

        if comp3_score > 0:
            total_score += comp3_score

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ------------------------------------------------------------------
    # Component 4: Dark blue header rows 3, 7, 14 (bold, fill, white text) (0.20 pts)
    # Initial file has no fill and no bold on header rows
    # Each correctly formatted header row earns 0.067 pts (3 rows * 0.067 = 0.20)
    # ------------------------------------------------------------------
    try:
        comp4_details = []
        header_rows = [3, 7, 14]
        header_row_passed = 0

        for row_num in header_rows:
            # Check first column of each header row for formatting
            row_issues = []
            cell = ws.cell(row=row_num, column=1)

            # Bold check
            if not cell.font.bold:
                row_issues.append(f"col1 not bold")

            # Dark blue fill check
            fill = get_fill_color_rgb(cell)
            if not (fill and fill.upper() == DARK_BLUE_FILL.upper()):
                row_issues.append(f"col1 fill={fill}")

            # White font color check
            fc = get_font_color_rgb(cell)
            if not (fc and fc.upper() in (WHITE_FONT.upper(), 'FFFFFF')):
                row_issues.append(f"col1 fontcolor={fc}")

            if len(row_issues) == 0:
                header_row_passed += 1
                comp4_details.append(f"row{row_num} OK")
            else:
                comp4_details.append(f"row{row_num} FAIL ({'; '.join(row_issues)})")

        # 0.20 total / 3 rows = ~0.0667 per row
        comp4_score = round(header_row_passed * (0.20 / 3), 4)
        if comp4_score >= 0.18:
            print(f"PASS: Component 4 — Dark blue header rows ({comp4_score:.2f} pts) — {'; '.join(comp4_details)}")
        else:
            print(f"PARTIAL: Component 4 — Dark blue header rows ({comp4_score:.2f}/0.20 pts) — {'; '.join(comp4_details)}")

        if comp4_score > 0:
            total_score += comp4_score

    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ------------------------------------------------------------------
    # Component 5: Bar chart embedded for regional data (0.10 pts)
    # Initial file has no charts
    # ------------------------------------------------------------------
    try:
        charts = ws._charts
        num_charts = len(charts)
        if num_charts >= 1:
            chart = charts[0]
            # Verify it is a bar chart type
            chart_type_name = type(chart).__name__
            if 'Bar' in chart_type_name:
                print(f"PASS: Component 5 — Bar chart found ({num_charts} chart(s), type={chart_type_name}) (0.10 pts)")
                total_score += 0.10
            elif chart_type_name:
                # Still award partial for any chart (agent attempted but used wrong type)
                print(f"PARTIAL: Component 5 — Chart found but type={chart_type_name}, expected BarChart (0.05 pts)")
                total_score += 0.05
        else:
            print(f"FAIL: Component 5 — No charts found in ExecReport sheet (expected >=1 bar chart)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # ------------------------------------------------------------------
    # Component 6: Print header/footer configured (0.10 pts)
    # Initial file has empty header/footer
    # ------------------------------------------------------------------
    try:
        comp6_score = 0.0
        comp6_details = []

        # Sub-check 6a: oddHeader has content (0.06 pts)
        header_str = str(ws.oddHeader) if ws.oddHeader else ''
        if header_str and len(header_str.strip()) > 0:
            comp6_score += 0.06
            comp6_details.append(f"header set ('{header_str[:60]}')")
        else:
            comp6_details.append("FAIL: header empty")

        # Sub-check 6b: footer has page number (0.04 pts)
        footer_str = str(ws.oddFooter) if ws.oddFooter else ''
        if footer_str and ('&P' in footer_str or 'Page' in footer_str):
            comp6_score += 0.04
            comp6_details.append(f"footer has page number ('{footer_str[:60]}')")
        elif footer_str and len(footer_str.strip()) > 0:
            comp6_score += 0.02
            comp6_details.append(f"footer set but no page number ('{footer_str[:60]}')")
        else:
            comp6_details.append("FAIL: footer empty")

        if comp6_score >= 0.08:
            print(f"PASS: Component 6 — Print header/footer ({comp6_score:.2f} pts) — {'; '.join(comp6_details)}")
        else:
            print(f"PARTIAL: Component 6 — Print header/footer ({comp6_score:.2f}/0.10 pts) — {'; '.join(comp6_details)}")

        if comp6_score > 0:
            total_score += comp6_score

    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score:.4f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
