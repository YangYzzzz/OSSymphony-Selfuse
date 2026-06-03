"""
Initial Setup: Change tab color of quarterly sheets to green
Task ID: calc_gsi_034
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_034'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # Common styles
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    currency_fmt = '$#,##0'
    pct_fmt = '0.0%'

    # --- Q1 Sheet ---
    ws_q1 = wb.active
    ws_q1.title = 'Q1'
    q1_headers = ['Category', 'January', 'February', 'March', 'Q1 Total']
    for c, h in enumerate(q1_headers, 1):
        cell = ws_q1.cell(row=1, column=c, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
    q1_data = [
        ['Revenue', 125000, 138000, 142000],
        ['Cost of Goods Sold', 52000, 57500, 59000],
        ['Gross Profit', 73000, 80500, 83000],
        ['Marketing', 15000, 16200, 14800],
        ['Salaries', 35000, 35000, 35000],
        ['Rent & Utilities', 8500, 8500, 8500],
        ['Software Licenses', 3200, 3200, 3200],
        ['Travel & Entertainment', 4500, 5100, 3800],
        ['Total Operating Expenses', 66200, 68000, 65300],
        ['Operating Income', 6800, 12500, 17700],
        ['Tax (25%)', 1700, 3125, 4425],
        ['Net Income', 5100, 9375, 13275],
    ]
    for r, row_data in enumerate(q1_data, 2):
        ws_q1.cell(row=r, column=1, value=row_data[0])
        for c, val in enumerate(row_data[1:], 2):
            cell = ws_q1.cell(row=r, column=c, value=val)
            cell.number_format = currency_fmt
        # Q1 Total formula
        ws_q1.cell(row=r, column=5, value=f'=SUM(B{r}:D{r})').number_format = currency_fmt
    ws_q1.column_dimensions['A'].width = 25
    for col_letter in ['B', 'C', 'D', 'E']:
        ws_q1.column_dimensions[col_letter].width = 15

    # --- Q2 Sheet ---
    ws_q2 = wb.create_sheet('Q2')
    q2_headers = ['Category', 'April', 'May', 'June', 'Q2 Total']
    for c, h in enumerate(q2_headers, 1):
        cell = ws_q2.cell(row=1, column=c, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
    q2_data = [
        ['Revenue', 148000, 155000, 162000],
        ['Cost of Goods Sold', 61500, 64000, 67000],
        ['Gross Profit', 86500, 91000, 95000],
        ['Marketing', 18000, 17500, 19200],
        ['Salaries', 36000, 36000, 36000],
        ['Rent & Utilities', 8500, 8500, 8500],
        ['Software Licenses', 3200, 3200, 3200],
        ['Travel & Entertainment', 5200, 6300, 4100],
        ['Total Operating Expenses', 70900, 71500, 71000],
        ['Operating Income', 15600, 19500, 24000],
        ['Tax (25%)', 3900, 4875, 6000],
        ['Net Income', 11700, 14625, 18000],
    ]
    for r, row_data in enumerate(q2_data, 2):
        ws_q2.cell(row=r, column=1, value=row_data[0])
        for c, val in enumerate(row_data[1:], 2):
            cell = ws_q2.cell(row=r, column=c, value=val)
            cell.number_format = currency_fmt
        ws_q2.cell(row=r, column=5, value=f'=SUM(B{r}:D{r})').number_format = currency_fmt
    ws_q2.column_dimensions['A'].width = 25
    for col_letter in ['B', 'C', 'D', 'E']:
        ws_q2.column_dimensions[col_letter].width = 15

    # --- Q3 Sheet ---
    ws_q3 = wb.create_sheet('Q3')
    q3_headers = ['Category', 'July', 'August', 'September', 'Q3 Total']
    for c, h in enumerate(q3_headers, 1):
        cell = ws_q3.cell(row=1, column=c, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
    q3_data = [
        ['Revenue', 158000, 149000, 165000],
        ['Cost of Goods Sold', 65500, 62000, 68500],
        ['Gross Profit', 92500, 87000, 96500],
        ['Marketing', 20000, 15000, 18500],
        ['Salaries', 37000, 37000, 37000],
        ['Rent & Utilities', 9000, 9000, 9000],
        ['Software Licenses', 3500, 3500, 3500],
        ['Travel & Entertainment', 3800, 7200, 5400],
        ['Total Operating Expenses', 73300, 71700, 73400],
        ['Operating Income', 19200, 15300, 23100],
        ['Tax (25%)', 4800, 3825, 5775],
        ['Net Income', 14400, 11475, 17325],
    ]
    for r, row_data in enumerate(q3_data, 2):
        ws_q3.cell(row=r, column=1, value=row_data[0])
        for c, val in enumerate(row_data[1:], 2):
            cell = ws_q3.cell(row=r, column=c, value=val)
            cell.number_format = currency_fmt
        ws_q3.cell(row=r, column=5, value=f'=SUM(B{r}:D{r})').number_format = currency_fmt
    ws_q3.column_dimensions['A'].width = 25
    for col_letter in ['B', 'C', 'D', 'E']:
        ws_q3.column_dimensions[col_letter].width = 15

    # --- Q4 Sheet ---
    ws_q4 = wb.create_sheet('Q4')
    q4_headers = ['Category', 'October', 'November', 'December', 'Q4 Total']
    for c, h in enumerate(q4_headers, 1):
        cell = ws_q4.cell(row=1, column=c, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
    q4_data = [
        ['Revenue', 172000, 185000, 198000],
        ['Cost of Goods Sold', 71000, 76500, 82000],
        ['Gross Profit', 101000, 108500, 116000],
        ['Marketing', 22000, 25000, 28000],
        ['Salaries', 38000, 38000, 38000],
        ['Rent & Utilities', 9000, 9000, 9000],
        ['Software Licenses', 3500, 3500, 3500],
        ['Travel & Entertainment', 6500, 4200, 8800],
        ['Total Operating Expenses', 79000, 79700, 87300],
        ['Operating Income', 22000, 28800, 28700],
        ['Tax (25%)', 5500, 7200, 7175],
        ['Net Income', 16500, 21600, 21525],
    ]
    for r, row_data in enumerate(q4_data, 2):
        ws_q4.cell(row=r, column=1, value=row_data[0])
        for c, val in enumerate(row_data[1:], 2):
            cell = ws_q4.cell(row=r, column=c, value=val)
            cell.number_format = currency_fmt
        ws_q4.cell(row=r, column=5, value=f'=SUM(B{r}:D{r})').number_format = currency_fmt
    ws_q4.column_dimensions['A'].width = 25
    for col_letter in ['B', 'C', 'D', 'E']:
        ws_q4.column_dimensions[col_letter].width = 15

    # --- Annual Summary Sheet ---
    ws_annual = wb.create_sheet('Annual Summary')
    annual_headers = ['Metric', 'Q1', 'Q2', 'Q3', 'Q4', 'Annual Total']
    for c, h in enumerate(annual_headers, 1):
        cell = ws_annual.cell(row=1, column=c, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
        cell.alignment = header_align
    annual_data = [
        ['Total Revenue', 405000, 465000, 472000, 555000],
        ['Total COGS', 168500, 192500, 196000, 229500],
        ['Gross Profit', 236500, 272500, 276000, 325500],
        ['Total OpEx', 199500, 213400, 218400, 246000],
        ['Operating Income', 37000, 59100, 57600, 79500],
        ['Total Tax', 9250, 14775, 14400, 19875],
        ['Net Income', 27750, 44325, 43200, 59625],
    ]
    for r, row_data in enumerate(annual_data, 2):
        ws_annual.cell(row=r, column=1, value=row_data[0])
        for c, val in enumerate(row_data[1:], 2):
            cell = ws_annual.cell(row=r, column=c, value=val)
            cell.number_format = currency_fmt
        ws_annual.cell(row=r, column=6, value=f'=SUM(B{r}:E{r})').number_format = currency_fmt
    ws_annual.column_dimensions['A'].width = 22
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws_annual.column_dimensions[col_letter].width = 16

    # --- Charts Sheet ---
    ws_charts = wb.create_sheet('Charts')
    ws_charts.cell(row=1, column=1, value='Revenue & Profit Trends')
    ws_charts.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True)
    ws_charts.cell(row=3, column=1, value='Quarter')
    ws_charts.cell(row=3, column=2, value='Revenue')
    ws_charts.cell(row=3, column=3, value='Net Income')
    chart_data = [
        ['Q1', 405000, 27750],
        ['Q2', 465000, 44325],
        ['Q3', 472000, 43200],
        ['Q4', 555000, 59625],
    ]
    for r, row_data in enumerate(chart_data, 4):
        for c, val in enumerate(row_data, 1):
            ws_charts.cell(row=r, column=c, value=val)

    from openpyxl.chart import BarChart, Reference
    chart = BarChart()
    chart.type = "col"
    chart.title = "Quarterly Revenue vs Net Income"
    chart.y_axis.title = "Amount ($)"
    chart.x_axis.title = "Quarter"
    data_ref = Reference(ws_charts, min_col=2, min_row=3, max_col=3, max_row=7)
    cats_ref = Reference(ws_charts, min_col=1, min_row=4, max_row=7)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width = 20
    chart.height = 12
    ws_charts.add_chart(chart, "A9")

    # --- Assumptions Sheet ---
    ws_assumptions = wb.create_sheet('Assumptions')
    ws_assumptions.cell(row=1, column=1, value='Key Assumptions').font = Font(size=14, bold=True)
    assumptions = [
        ['Tax Rate', '25%'],
        ['Annual Salary Increase', '3% per quarter adjustment'],
        ['Marketing Budget Growth', '5-10% quarterly increase'],
        ['COGS Ratio', '~41% of revenue'],
        ['Rent Escalation', 'Step increase in Q3 (+$500/mo)'],
        ['Software License Renewal', 'Annual renewal in Q3'],
        ['Revenue Growth Target', '8-12% QoQ'],
        ['Inflation Assumption', '3.2% annual'],
        ['Headcount', '42 FTEs across departments'],
        ['Office Locations', 'San Francisco (HQ), Austin (satellite)'],
    ]
    ws_assumptions.cell(row=3, column=1, value='Parameter').font = Font(bold=True)
    ws_assumptions.cell(row=3, column=2, value='Value / Note').font = Font(bold=True)
    for r, row_data in enumerate(assumptions, 4):
        ws_assumptions.cell(row=r, column=1, value=row_data[0])
        ws_assumptions.cell(row=r, column=2, value=row_data[1])
    ws_assumptions.column_dimensions['A'].width = 28
    ws_assumptions.column_dimensions['B'].width = 40

    # NO tab colors set on any sheet (that's the task for the agent)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
