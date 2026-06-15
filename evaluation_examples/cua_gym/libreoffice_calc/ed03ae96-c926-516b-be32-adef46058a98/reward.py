"""
Reward Script: Change tab color of Q1-Q4 sheets to green
Task ID: calc_gsi_034
Domain: libreoffice_calc
Scoring:
  - Component 1: Q1 sheet tab color is green (0.25 pts)
  - Component 2: Q2 sheet tab color is green (0.25 pts)
  - Component 3: Q3 sheet tab color is green (0.25 pts)
  - Component 4: Q4 sheet tab color is green (0.25 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_034'


def is_green_color(color_obj):
    """
    Check if an openpyxl Color object represents a shade of green.
    Accepts common green shades by checking the RGB hex.
    Green colors have high G channel relative to R and B.
    """
    if color_obj is None:
        return False
    try:
        rgb_str = color_obj.rgb
        if rgb_str is None:
            return False
        # rgb_str is ARGB format like '0000B050' or 'FF00B050'
        # Extract the last 6 chars for RGB
        rgb_hex = rgb_str[-6:]
        r = int(rgb_hex[0:2], 16)
        g = int(rgb_hex[2:4], 16)
        b = int(rgb_hex[4:6], 16)
        # Green: G channel is dominant (significantly higher than R and B)
        if g >= 128 and g > r * 2 and g > b * 2:
            return True
        return False
    except Exception:
        return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    quarter_sheets = ['Q1', 'Q2', 'Q3', 'Q4']

    # Precondition: all four quarter sheets must exist
    for name in quarter_sheets:
        if name not in wb.sheetnames:
            print(f"CRITICAL: Sheet '{name}' not found in workbook. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0

    # Component 1: Q1 sheet tab color is green (0.25 points)
    try:
        ws = wb['Q1']
        tab_color = ws.sheet_properties.tabColor
        if is_green_color(tab_color):
            print(f"PASS: Component 1 — Q1 tab color is green (rgb={tab_color.rgb}) (0.25 pts)")
            total_score += 0.25
        else:
            color_info = tab_color.rgb if tab_color else 'None'
            print(f"FAIL: Component 1 — Q1 tab color is not green, found: {color_info}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Q2 sheet tab color is green (0.25 points)
    try:
        ws = wb['Q2']
        tab_color = ws.sheet_properties.tabColor
        if is_green_color(tab_color):
            print(f"PASS: Component 2 — Q2 tab color is green (rgb={tab_color.rgb}) (0.25 pts)")
            total_score += 0.25
        else:
            color_info = tab_color.rgb if tab_color else 'None'
            print(f"FAIL: Component 2 — Q2 tab color is not green, found: {color_info}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Q3 sheet tab color is green (0.25 points)
    try:
        ws = wb['Q3']
        tab_color = ws.sheet_properties.tabColor
        if is_green_color(tab_color):
            print(f"PASS: Component 3 — Q3 tab color is green (rgb={tab_color.rgb}) (0.25 pts)")
            total_score += 0.25
        else:
            color_info = tab_color.rgb if tab_color else 'None'
            print(f"FAIL: Component 3 — Q3 tab color is not green, found: {color_info}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Q4 sheet tab color is green (0.25 points)
    try:
        ws = wb['Q4']
        tab_color = ws.sheet_properties.tabColor
        if is_green_color(tab_color):
            print(f"PASS: Component 4 — Q4 tab color is green (rgb={tab_color.rgb}) (0.25 pts)")
            total_score += 0.25
        else:
            color_info = tab_color.rgb if tab_color else 'None'
            print(f"FAIL: Component 4 — Q4 tab color is not green, found: {color_info}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
