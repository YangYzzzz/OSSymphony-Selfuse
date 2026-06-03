"""
Initial Setup: Create multi-year financial data for pivot table task
Task ID: calc_pivot_094
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_094'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def generate_data():
    """Generate 600 rows: 150 per year, 50 per category per year.
    Controlled so Revenue growth 2023 vs 2022 is ~12%.
    """
    random.seed(42)
    categories = ['Revenue', 'COGS', 'OpEx']
    years = [2021, 2022, 2023, 2024]

    # Target yearly totals (sum of 50 entries each):
    # Revenue: 2021=540k, 2022=600k, 2023=672k (+12%), 2024=725k (+7.9%)
    # COGS:    2021=264k, 2022=288k (+9.1%), 2023=318k (+10.4%), 2024=336k (+5.7%)
    # OpEx:    2021=180k, 2022=192k (+6.7%), 2023=210k (+9.4%), 2024=228k (+8.6%)
    target_means = {
        'Revenue': {2021: 10800, 2022: 12000, 2023: 12960, 2024: 14500},
        'COGS':    {2021: 5280,  2022: 5760,  2023: 6360,  2024: 6720},
        'OpEx':    {2021: 3600,  2022: 3840,  2023: 4200,  2024: 4560},
    }
    # Each category has 50 entries per year, mean = target_means[cat][year]

    data_rows = []
    row_id = 1
    for year in years:
        for cat in categories:
            mean_val = target_means[cat][year]
            for i in range(50):
                month = (i % 12) + 1
                # Add realistic variation: +/- 20% noise
                noise = random.uniform(0.80, 1.20)
                amount = round(mean_val * noise, 2)
                date_str = f'{month:02d}/{year}'
                data_rows.append([row_id, date_str, cat, amount])
                row_id += 1

    # Shuffle to make it look natural (not grouped by year/category)
    random.shuffle(data_rows)
    # Re-assign sequential IDs after shuffle
    for i, row in enumerate(data_rows):
        row[0] = i + 1

    return data_rows

def create_initial():
    data_rows = generate_data()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'FinancialHistory'

    # --- Headers ---
    headers = ['ID', 'Date', 'Category', 'Amount']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # --- Write Data ---
    for r, row_data in enumerate(data_rows, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if c == 4:
                cell.number_format = '#,##0.00'
            elif c == 1:
                cell.alignment = Alignment(horizontal='center')

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Total data rows: {len(data_rows)}')

    # Verify totals
    from collections import defaultdict
    sums = defaultdict(lambda: defaultdict(float))
    for row_data in data_rows:
        date_str = row_data[1]
        cat = row_data[2]
        amount = row_data[3]
        year = int(date_str.split('/')[1])
        sums[year][cat] += amount
    for year in sorted(sums.keys()):
        print(f"  {year}: " + ", ".join(f"{cat}={sums[year][cat]:,.2f}" for cat in ['Revenue', 'COGS', 'OpEx']))
    # Check Revenue growth
    rev_22 = sums[2022]['Revenue']
    rev_23 = sums[2023]['Revenue']
    print(f"Revenue growth 2023 vs 2022: {(rev_23 - rev_22)/rev_22*100:.2f}%")

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
