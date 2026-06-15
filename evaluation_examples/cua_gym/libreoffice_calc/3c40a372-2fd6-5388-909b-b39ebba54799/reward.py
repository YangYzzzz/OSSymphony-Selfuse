"""
Reward Script: Create a pivot table showing year-over-year growth percentage
Task ID: calc_pivot_094
Domain: libreoffice_calc
Scoring:
  Component 1: PivotTable sheet exists (0.20)
  Component 2: Correct structure — Year rows + Category columns (0.25)
  Component 3: Base year (2021) has blank values, other years have numeric % values (0.25)
  Component 4: Percentage values are approximately correct (0.20)
  Component 5: Percentage number format applied on data cells (0.10)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_094'

# Expected approximate values from golden (% difference from previous year)
# Row = year, columns = COGS, OpEx, Revenue
EXPECTED_VALUES = {
    2022: {'COGS': 0.0979, 'OpEx': 0.0676, 'Revenue': 0.1254},
    2023: {'COGS': 0.0948, 'OpEx': 0.1005, 'Revenue': 0.1199},
    2024: {'COGS': 0.0359, 'OpEx': 0.1033, 'Revenue': 0.0717},
}

TOLERANCE = 0.02  # Allow 2% absolute tolerance on percentage values


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: PivotTable sheet exists (0.20 points)
    # This is the primary task-introduced change — initial has only FinancialHistory
    try:
        pivot_sheet_names = [s for s in wb.sheetnames if s != 'FinancialHistory']
        if len(pivot_sheet_names) >= 1:
            # Find a sheet that looks like a pivot table (has year data)
            pivot_ws = None
            for sn in pivot_sheet_names:
                ws_candidate = wb[sn]
                # Check if it has year values somewhere in column A
                has_years = False
                for row in ws_candidate.iter_rows(min_col=1, max_col=1, values_only=True):
                    if row[0] in (2021, 2022, 2023, 2024):
                        has_years = True
                        break
                if has_years:
                    pivot_ws = ws_candidate
                    break
            if pivot_ws is not None:
                print(f"PASS: Component 1 — Pivot table sheet found: '{pivot_ws.title}' (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 1 — Extra sheets exist but none contain year data: {pivot_sheet_names}")
        else:
            print(f"FAIL: Component 1 — No pivot table sheet found. Sheets: {wb.sheetnames}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    if total_score < 0.20:
        # No pivot table sheet found, cannot continue
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # From here on, pivot_ws is the pivot table sheet
    # Find the header row and data layout
    # Expected: a header row with Year + category names, then data rows with years

    # Component 2: Correct structure — Year rows + Category columns (0.25 points)
    try:
        # Find the header row (contains 'Year' or year-like label + category names)
        header_row = None
        year_col = None
        category_cols = {}  # maps category name -> column index

        for r in range(1, min(pivot_ws.max_row + 1, 10)):
            row_vals = []
            for c in range(1, min(pivot_ws.max_column + 1, 10)):
                row_vals.append((c, pivot_ws.cell(row=r, column=c).value))

            # Check if this row has 'Year' and category names
            row_texts = [str(v[1]).strip().lower() if v[1] is not None else '' for v in row_vals]
            if 'year' in row_texts:
                header_row = r
                for col_idx, val in row_vals:
                    if val is not None:
                        val_str = str(val).strip()
                        val_lower = val_str.lower()
                        if val_lower == 'year':
                            year_col = col_idx
                        elif val_lower in ('cogs', 'opex', 'revenue'):
                            category_cols[val_str] = col_idx
                break

        structure_score = 0.0
        if header_row is not None and year_col is not None:
            structure_score += 0.10  # Found header row with Year column
            print(f"  Found header row at row {header_row}, Year in column {year_col}")
        else:
            print(f"FAIL: Component 2 — Could not find header row with 'Year' column")

        # Check for at least 2 category columns
        if len(category_cols) >= 2:
            structure_score += 0.05
            print(f"  Found {len(category_cols)} category columns: {list(category_cols.keys())}")
        else:
            print(f"FAIL: Component 2 — Expected at least 2 category columns, found: {list(category_cols.keys())}")

        # Check for year rows (at least years 2021-2024)
        if header_row is not None and year_col is not None:
            year_rows = {}  # year -> row number
            for r in range(header_row + 1, min(pivot_ws.max_row + 1, header_row + 10)):
                val = pivot_ws.cell(row=r, column=year_col).value
                if isinstance(val, (int, float)) and 2021 <= val <= 2024:
                    year_rows[int(val)] = r

            if len(year_rows) >= 4:
                structure_score += 0.10
                print(f"  Found year rows: {sorted(year_rows.keys())}")
            elif len(year_rows) >= 2:
                structure_score += 0.05
                print(f"  Found partial year rows: {sorted(year_rows.keys())} (expected 4)")
            else:
                print(f"FAIL: Component 2 — Found {len(year_rows)} year rows, expected 4")

        if structure_score > 0:
            print(f"PASS: Component 2 — Structure verified ({structure_score:.2f} pts)")
            total_score += structure_score
        else:
            print(f"FAIL: Component 2 — Structure not correct")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Base year (2021) has blank/None values, other years have numeric values (0.25 points)
    try:
        comp3_score = 0.0

        if header_row is not None and year_col is not None and len(category_cols) >= 1 and len(year_rows) >= 1:
            # Check 2021 base year: values should be None/blank/0
            if 2021 in year_rows:
                base_row = year_rows[2021]
                all_blank = True
                for cat_name, cat_col in category_cols.items():
                    val = pivot_ws.cell(row=base_row, column=cat_col).value
                    if val is not None and val != '' and val != 0:
                        all_blank = False
                        break
                if all_blank:
                    comp3_score += 0.10
                    print(f"  2021 base year has blank values (correct)")
                else:
                    print(f"  FAIL: 2021 should have blank values but has data")

            # Check that 2022-2024 have numeric percentage values
            years_with_data = 0
            for year in [2022, 2023, 2024]:
                if year in year_rows:
                    row_num = year_rows[year]
                    has_numeric = False
                    for cat_name, cat_col in category_cols.items():
                        val = pivot_ws.cell(row=row_num, column=cat_col).value
                        if isinstance(val, (int, float)) and val != 0:
                            has_numeric = True
                            break
                    if has_numeric:
                        years_with_data += 1

            if years_with_data >= 3:
                comp3_score += 0.15
                print(f"  Years 2022-2024 all have numeric data (correct)")
            elif years_with_data >= 1:
                comp3_score += 0.05 * years_with_data
                print(f"  {years_with_data}/3 years have numeric data")
            else:
                print(f"  FAIL: No years 2022-2024 have numeric data")

            if comp3_score > 0:
                print(f"PASS: Component 3 — Base year blank + others numeric ({comp3_score:.2f} pts)")
                total_score += comp3_score
            else:
                print(f"FAIL: Component 3 — Data pattern not correct")
        else:
            print(f"FAIL: Component 3 — Cannot check (missing structure from Component 2)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Percentage values are approximately correct (0.20 points)
    try:
        comp4_score = 0.0
        matched_cells = 0
        total_cells = 0

        if header_row is not None and len(category_cols) >= 1 and len(year_rows) >= 1:
            for year, expected_cats in EXPECTED_VALUES.items():
                if year not in year_rows:
                    continue
                row_num = year_rows[year]
                for cat_name, expected_val in expected_cats.items():
                    # Find matching category column (case-insensitive)
                    matched_col = None
                    for cn, ci in category_cols.items():
                        if cn.lower() == cat_name.lower():
                            matched_col = ci
                            break
                    if matched_col is None:
                        continue

                    total_cells += 1
                    actual = pivot_ws.cell(row=row_num, column=matched_col).value
                    if isinstance(actual, (int, float)):
                        if abs(actual - expected_val) <= TOLERANCE:
                            matched_cells += 1
                        else:
                            print(f"  Value mismatch: {year}/{cat_name} expected ~{expected_val:.4f}, got {actual:.4f}")

            if total_cells > 0:
                ratio = matched_cells / total_cells
                comp4_score = round(0.20 * ratio, 4)
                print(f"  Matched {matched_cells}/{total_cells} cells (ratio={ratio:.2f})")
            else:
                print(f"  No cells to compare")

            if comp4_score > 0:
                print(f"PASS: Component 4 — Values approximately correct ({comp4_score:.2f} pts)")
                total_score += comp4_score
            else:
                print(f"FAIL: Component 4 — Values not matching expected")
        else:
            print(f"FAIL: Component 4 — Cannot check (missing structure)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Percentage number format on data cells (0.10 points)
    try:
        comp5_score = 0.0
        pct_format_count = 0
        data_cell_count = 0

        if header_row is not None and len(category_cols) >= 1 and len(year_rows) >= 1:
            for year in [2022, 2023, 2024]:
                if year not in year_rows:
                    continue
                row_num = year_rows[year]
                for cat_name, cat_col in category_cols.items():
                    cell = pivot_ws.cell(row=row_num, column=cat_col)
                    if isinstance(cell.value, (int, float)):
                        data_cell_count += 1
                        nf = cell.number_format
                        if nf is not None and '%' in str(nf):
                            pct_format_count += 1

            if data_cell_count > 0 and pct_format_count >= data_cell_count * 0.5:
                comp5_score = 0.10
                print(f"  {pct_format_count}/{data_cell_count} data cells have % format")
            elif pct_format_count > 0:
                comp5_score = 0.05
                print(f"  Partial: {pct_format_count}/{data_cell_count} data cells have % format")
            else:
                print(f"  FAIL: No data cells have percentage number format")

            if comp5_score > 0:
                print(f"PASS: Component 5 — Percentage format applied ({comp5_score:.2f} pts)")
                total_score += comp5_score
            else:
                print(f"FAIL: Component 5 — Percentage format not applied")
        else:
            print(f"FAIL: Component 5 — Cannot check (missing structure)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
