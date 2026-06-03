"""
Reward Script: Create a sales leaderboard with data bars and rank formatting.
Task ID: calc_gpm_017
Domain: libreoffice_calc
Scoring:
  Component 1: Title row merged & styled (0.20)
  Component 2: Header row bold, centered, light green fill (0.15)
  Component 3: Gold/silver/bronze rank cell fills (0.15)
  Component 4: Currency formatting on C3:D10 (0.10)
  Component 5: Attainment formulas (=C/D) with percentage format in E3:E10 (0.15)
  Component 6: Conditional formatting rules on C3:C10 and E3:E10 (0.15)
  Component 7: Borders on A2:E10 (0.10)
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_017'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Ensure 'Leaderboard' sheet exists
    if 'Leaderboard' not in wb.sheetnames:
        print("FAIL: 'Leaderboard' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Leaderboard']

    # ============================================================
    # Component 1: Title row A1:E1 merged, bold, 14pt, centered,
    #   dark green fill, white font (0.20 points)
    # ============================================================
    try:
        # Check merge
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        has_merge = any('A1' in r and 'E1' in r for r in merged_ranges)

        cell_a1 = ws['A1']
        is_bold = cell_a1.font.bold is True
        is_14pt = cell_a1.font.size is not None and abs(cell_a1.font.size - 14) < 0.5
        is_centered = cell_a1.alignment.horizontal == 'center'

        # Check dark green fill (006400)
        try:
            fill_rgb = cell_a1.fill.fgColor.rgb
            has_green_fill = fill_rgb is not None and '006400' in str(fill_rgb).upper()
        except Exception:
            has_green_fill = False

        # Check white font color
        try:
            font_rgb = cell_a1.font.color.rgb
            has_white_font = font_rgb is not None and 'FFFFFF' in str(font_rgb).upper()
        except Exception:
            has_white_font = False

        # Sub-scoring: merge + bold/size + centered + fill + font color
        sub = 0
        if has_merge:
            sub += 1
        if is_bold and is_14pt:
            sub += 1
        if is_centered:
            sub += 1
        if has_green_fill:
            sub += 1
        if has_white_font:
            sub += 1

        comp1 = 0.20 * (sub / 5)
        if comp1 > 0:
            print(f"PASS: Component 1 -- Title styling ({comp1:.3f} pts, {sub}/5 sub-checks)")
            total_score += comp1
        else:
            print(f"FAIL: Component 1 -- Title styling: merge={has_merge}, bold={is_bold}, 14pt={is_14pt}, centered={is_centered}, green_fill={has_green_fill}, white_font={has_white_font}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # ============================================================
    # Component 2: Header row 2 bold, centered, light green fill (0.15 points)
    # ============================================================
    try:
        header_cells = ['A2', 'B2', 'C2', 'D2', 'E2']
        bold_count = 0
        centered_count = 0
        green_fill_count = 0

        for coord in header_cells:
            cell = ws[coord]
            if cell.font.bold is True:
                bold_count += 1
            if cell.alignment.horizontal == 'center':
                centered_count += 1
            try:
                fill_rgb = cell.fill.fgColor.rgb
                # Light green fill: 90EE90
                if fill_rgb is not None and '90EE90' in str(fill_rgb).upper():
                    green_fill_count += 1
            except Exception:
                pass

        # Need majority (>=4 out of 5) for each sub-check
        sub_bold = bold_count >= 4
        sub_centered = centered_count >= 4
        sub_green = green_fill_count >= 4

        sub = sum([sub_bold, sub_centered, sub_green])
        comp2 = 0.15 * (sub / 3)
        if comp2 > 0:
            print(f"PASS: Component 2 -- Header row ({comp2:.3f} pts, bold={bold_count}/5, centered={centered_count}/5, green={green_fill_count}/5)")
            total_score += comp2
        else:
            print(f"FAIL: Component 2 -- Header row: bold={bold_count}/5, centered={centered_count}/5, green_fill={green_fill_count}/5")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # ============================================================
    # Component 3: Gold/Silver/Bronze fills on rank cells A3, A4, A5 (0.15 points)
    # ============================================================
    try:
        medal_checks = {
            'A3': 'FFD700',  # gold
            'A4': 'C0C0C0',  # silver
            'A5': 'CD7F32',  # bronze
        }
        medal_pass = 0
        for coord, expected_hex in medal_checks.items():
            cell = ws[coord]
            try:
                fill_rgb = str(cell.fill.fgColor.rgb).upper()
                if expected_hex in fill_rgb:
                    medal_pass += 1
                else:
                    print(f"  INFO: {coord} fill={fill_rgb}, expected contains {expected_hex}")
            except Exception:
                print(f"  INFO: {coord} fill could not be read")

        comp3 = 0.15 * (medal_pass / 3)
        if comp3 > 0:
            print(f"PASS: Component 3 -- Medal fills ({comp3:.3f} pts, {medal_pass}/3)")
            total_score += comp3
        else:
            print(f"FAIL: Component 3 -- Medal fills: {medal_pass}/3 matched")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # ============================================================
    # Component 4: Currency formatting ($#,##0) on C3:D10 (0.10 points)
    # ============================================================
    try:
        currency_count = 0
        total_cells = 0
        for row in range(3, 11):
            for col in [3, 4]:  # C, D
                total_cells += 1
                cell = ws.cell(row=row, column=col)
                nf = cell.number_format
                if nf and '$' in str(nf):
                    currency_count += 1

        # Need at least 12 out of 16 cells formatted
        if currency_count >= 12:
            print(f"PASS: Component 4 -- Currency format ({currency_count}/{total_cells} cells) (0.100 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 -- Currency format: {currency_count}/{total_cells} cells have '$' format")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # ============================================================
    # Component 5: Attainment formulas in E3:E10 (=C/D) with percentage format (0.15 points)
    # ============================================================
    try:
        formula_count = 0
        pct_format_count = 0
        for row in range(3, 11):
            cell = ws.cell(row=row, column=5)  # column E
            val = cell.value
            # Check for formula referencing C and D columns
            if isinstance(val, str) and val.startswith('=') and 'C' in val.upper() and 'D' in val.upper():
                formula_count += 1
            nf = cell.number_format
            if nf and '%' in str(nf):
                pct_format_count += 1

        sub = 0
        if formula_count >= 6:
            sub += 1
        if pct_format_count >= 6:
            sub += 1

        comp5 = 0.15 * (sub / 2)
        if comp5 > 0:
            print(f"PASS: Component 5 -- Attainment formulas ({comp5:.3f} pts, formulas={formula_count}/8, pct_fmt={pct_format_count}/8)")
            total_score += comp5
        else:
            print(f"FAIL: Component 5 -- Attainment: formulas={formula_count}/8, pct_fmt={pct_format_count}/8")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    # ============================================================
    # Component 6: Conditional formatting (data bars on C3:C10 + rules on E3:E10) (0.15 points)
    # ============================================================
    try:
        databar_count = 0
        attainment_cf_count = 0

        for cf in ws.conditional_formatting:
            range_str = str(cf)
            for rule in cf.rules:
                if rule.type == 'dataBar':
                    # Data bar should cover C3:C10
                    if 'C3' in range_str or 'C' in range_str:
                        databar_count += 1
                elif rule.type == 'cellIs':
                    # Attainment conditional formatting on E column
                    if 'E3' in range_str or 'E' in range_str:
                        attainment_cf_count += 1

        sub = (1 if databar_count > 0 else 0) + (1 if attainment_cf_count > 0 else 0)
        comp6 = 0.15 * (sub / 2)
        if comp6 > 0:
            print(f"PASS: Component 6 -- Conditional formatting ({comp6:.3f} pts, databar={databar_count}, attainment_cf={attainment_cf_count})")
            total_score += comp6
        else:
            print(f"FAIL: Component 6 -- Conditional formatting: databar={databar_count}, attainment_cf={attainment_cf_count}")
    except Exception as e:
        print(f"ERROR: Component 6 -- {e}")

    # ============================================================
    # Component 7: Borders on A2:E10 (0.10 points)
    # ============================================================
    try:
        border_count = 0
        total_border_cells = 0
        for row in range(2, 11):
            for col in range(1, 6):  # A-E
                total_border_cells += 1
                cell = ws.cell(row=row, column=col)
                if isinstance(cell, MergedCell):
                    continue
                # Check if at least one border side has a style
                sides = [cell.border.left.style, cell.border.right.style,
                         cell.border.top.style, cell.border.bottom.style]
                if any(s is not None for s in sides):
                    border_count += 1

        # Need majority of cells to have borders
        if border_count >= 35:  # 45 total cells, need ~78%
            print(f"PASS: Component 7 -- Borders ({border_count}/{total_border_cells} cells) (0.100 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 7 -- Borders: {border_count}/{total_border_cells} cells have borders")
    except Exception as e:
        print(f"ERROR: Component 7 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.3f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
