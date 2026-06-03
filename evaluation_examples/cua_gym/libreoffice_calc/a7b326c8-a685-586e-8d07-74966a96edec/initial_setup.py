"""
Initial Setup: Create a sales leaderboard with data bars and rank formatting.
Task ID: calc_gpm_017
Domain: libreoffice_calc

Initial state: Plain data table with sales rep info. No formatting, no data bars,
no conditional formatting, no merged title, no borders, no currency format.
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_017'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leaderboard"

    # Title in A1 (plain, unmerged, unformatted)
    ws["A1"] = "Q1 Sales Leaderboard"

    # Headers in row 2 (plain, no formatting)
    headers = ["Rank", "Sales Rep", "Revenue", "Quota", "Attainment"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)

    # Data rows 3-10
    data = [
        [1, "Jessica", 285000, 250000],
        [2, "Marcus",  267000, 250000],
        [3, "Rachel",  248000, 250000],
        [4, "Kevin",   231000, 250000],
        [5, "Diana",   219000, 250000],
        [6, "Tyler",   198000, 250000],
        [7, "Priya",   182000, 250000],
        [8, "Alex",    165000, 250000],
    ]
    for r, row_data in enumerate(data, 3):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Attainment column (E) - plain values, no percentage format
    for r in range(3, 11):
        revenue = ws.cell(row=r, column=3).value
        quota = ws.cell(row=r, column=4).value
        ws.cell(row=r, column=5, value=revenue / quota)

    # Set reasonable column widths for readability
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
