"""
Initial Setup: Merge cells A1:D1 and center title in Sales sheet
Task ID: calc_gg1_006
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_006'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Sales ---
    ws = wb.active
    ws.title = 'Sales'

    # Row 1: Title in A1 only (NOT merged, NOT centered)
    ws.cell(row=1, column=1, value='Quarterly Sales Report')
    ws['A1'].font = Font(size=14, bold=True)
    # B1, C1, D1 left empty intentionally

    # Row 2: Column headers
    headers = ['Region', 'Q1 Sales', 'Q2 Sales', 'Q3 Sales']
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
        ws.cell(row=2, column=col).font = Font(bold=True)

    # Rows 3-14: Realistic sales data (12 regions)
    data = [
        ['Northeast', 145230, 162450, 158900],
        ['Southeast', 98750, 112300, 121500],
        ['Midwest', 87600, 94200, 103800],
        ['Southwest', 76400, 83100, 89700],
        ['Pacific Northwest', 112500, 128700, 135200],
        ['Mountain West', 54300, 61200, 67800],
        ['Mid-Atlantic', 134800, 141600, 149300],
        ['New England', 92100, 98400, 105600],
        ['Great Plains', 43200, 48700, 52100],
        ['Gulf Coast', 67800, 74500, 81200],
        ['Upper Midwest', 58900, 63400, 69700],
        ['Southern California', 156700, 168300, 175400],
    ]
    for r, row_data in enumerate(data, 3):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
