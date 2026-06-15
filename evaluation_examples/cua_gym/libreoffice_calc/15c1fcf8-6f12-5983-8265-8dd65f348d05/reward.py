"""
Reward Script: Merge cells A1:D1 and center 'Quarterly Sales Report' title
Task ID: calc_gg1_006
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): A1:D1 merge range exists
  Component 2 (0.3): A1 horizontal alignment is 'center'
  Component 3 (0.2): B1/C1/D1 are MergedCell objects AND A1 text is preserved
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_006'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Sales' sheet must exist
    if 'Sales' not in wb.sheetnames:
        print(f"CRITICAL: 'Sales' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Sales']

    # Component 1: A1:D1 merge range exists (0.5 points)
    # This is the core task requirement — cells A1 through D1 should be merged.
    try:
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        if 'A1:D1' in merged_ranges:
            print(f"PASS: Component 1 — A1:D1 merge range found (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — A1:D1 merge not found. Merged ranges: {merged_ranges}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: A1 horizontal alignment is 'center' (0.3 points)
    # The task requires the text to be centered in the merged cell.
    try:
        alignment = ws['A1'].alignment
        h_align = alignment.horizontal
        if h_align == 'center':
            print(f"PASS: Component 2 — A1 horizontal alignment is 'center' (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — A1 horizontal alignment is '{h_align}', expected 'center'")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: B1/C1/D1 are MergedCell objects AND A1 still has correct text (0.2 points)
    # Verifies the merge is structurally correct and text is preserved.
    try:
        b1_merged = isinstance(ws['B1'], MergedCell)
        c1_merged = isinstance(ws['C1'], MergedCell)
        d1_merged = isinstance(ws['D1'], MergedCell)
        a1_text = ws['A1'].value
        text_correct = (a1_text is not None and
                        str(a1_text).strip() == 'Quarterly Sales Report')

        if b1_merged and c1_merged and d1_merged and text_correct:
            print(f"PASS: Component 3 — B1/C1/D1 are merged cells, A1 text preserved (0.2 pts)")
            total_score += 0.2
        else:
            details = []
            if not b1_merged:
                details.append("B1 not merged")
            if not c1_merged:
                details.append("C1 not merged")
            if not d1_merged:
                details.append("D1 not merged")
            if not text_correct:
                details.append(f"A1 text is '{a1_text}', expected 'Quarterly Sales Report'")
            print(f"FAIL: Component 3 — {'; '.join(details)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
