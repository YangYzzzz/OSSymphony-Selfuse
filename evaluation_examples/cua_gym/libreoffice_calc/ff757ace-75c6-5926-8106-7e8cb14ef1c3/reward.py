"""
Reward Script: Create a formatted frequency distribution table from raw survey data.
Task ID: calc_gpm_024
Domain: libreoffice_calc
Scoring:
  C1: Headers (0.15), C2: Bins (0.10), C3: COUNTIFS (0.20), C4: Rel.Freq (0.15),
  C5: Cum.Freq (0.10), C6: Total row (0.10), C7: Header format (0.05),
  C8: Conditional formatting (0.10), C9: Thick border (0.05)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_024'

def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Survey' not in wb.sheetnames:
        print("FAIL: Sheet 'Survey' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Survey']

    # Component 1: Headers in C1:F1 with correct text and bold (0.15 points)
    try:
        expected_headers = {'C1': 'Bin', 'D1': 'Frequency', 'E1': 'Rel. Freq', 'F1': 'Cum. Freq'}
        headers_correct = 0
        headers_bold = 0
        for coord, expected in expected_headers.items():
            val = ws[coord].value
            if val is not None and str(val).strip() == expected:
                headers_correct += 1
            if ws[coord].font.bold:
                headers_bold += 1
        if headers_correct == 4 and headers_bold == 4:
            print(f"PASS: Component 1 - All 4 headers correct and bold (0.15 pts)")
            total_score += 0.15
        elif headers_correct >= 3:
            partial = 0.08
            print(f"PARTIAL: Component 1 - {headers_correct}/4 headers correct, {headers_bold}/4 bold ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - Headers: {headers_correct}/4 correct, {headers_bold}/4 bold")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Bin labels in C2:C7 (0.10 points)
    try:
        expected_bins = {
            'C2': '0-20', 'C3': '21-40', 'C4': '41-60',
            'C5': '61-80', 'C6': '81-100', 'C7': '101-120'
        }
        bins_correct = 0
        for coord, expected in expected_bins.items():
            val = ws[coord].value
            if val is not None and str(val).strip() == expected:
                bins_correct += 1
        if bins_correct == 6:
            print(f"PASS: Component 2 - All 6 bin labels correct (0.10 pts)")
            total_score += 0.10
        elif bins_correct >= 4:
            partial = 0.05
            print(f"PARTIAL: Component 2 - {bins_correct}/6 bins correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 - Only {bins_correct}/6 bin labels correct")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: COUNTIFS formulas in D2:D7 (0.20 points)
    try:
        countifs_count = 0
        for row in range(2, 8):
            val = ws.cell(row=row, column=4).value  # column D
            if val is not None and isinstance(val, str) and 'COUNTIF' in val.upper():
                countifs_count += 1
        if countifs_count == 6:
            print(f"PASS: Component 3 - All 6 COUNTIFS formulas present in D2:D7 (0.20 pts)")
            total_score += 0.20
        elif countifs_count >= 4:
            partial = 0.10
            print(f"PARTIAL: Component 3 - {countifs_count}/6 COUNTIFS formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 - Only {countifs_count}/6 COUNTIFS formulas in D2:D7")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Relative frequency formulas in E2:E7 with percentage format (0.15 points)
    try:
        rel_freq_formulas = 0
        pct_format_count = 0
        for row in range(2, 8):
            cell = ws.cell(row=row, column=5)  # column E
            val = cell.value
            if val is not None and isinstance(val, str) and '/' in val and 'SUM' in val.upper():
                rel_freq_formulas += 1
            # Also accept =D2/30 style or similar division formulas
            elif val is not None and isinstance(val, str) and val.startswith('=') and ('D' in val.upper()):
                rel_freq_formulas += 1
            nf = cell.number_format
            if nf is not None and '%' in str(nf):
                pct_format_count += 1

        if rel_freq_formulas >= 5 and pct_format_count >= 5:
            print(f"PASS: Component 4 - {rel_freq_formulas}/6 rel.freq formulas, {pct_format_count}/6 pct format (0.15 pts)")
            total_score += 0.15
        elif rel_freq_formulas >= 4:
            partial = 0.08
            print(f"PARTIAL: Component 4 - {rel_freq_formulas}/6 formulas, {pct_format_count}/6 pct format ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 - {rel_freq_formulas}/6 rel.freq formulas, {pct_format_count}/6 pct format")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Cumulative frequency formulas in F2:F7 (0.10 points)
    try:
        cum_freq_count = 0
        for row in range(2, 8):
            val = ws.cell(row=row, column=6).value  # column F
            if val is not None and isinstance(val, str) and val.startswith('='):
                cum_freq_count += 1
        if cum_freq_count >= 5:
            print(f"PASS: Component 5 - {cum_freq_count}/6 cumulative freq formulas in F2:F7 (0.10 pts)")
            total_score += 0.10
        elif cum_freq_count >= 3:
            partial = 0.05
            print(f"PARTIAL: Component 5 - {cum_freq_count}/6 cum.freq formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 - Only {cum_freq_count}/6 cumulative freq formulas in F2:F7")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    # Component 6: Total row - C9='Total' bold, D9=SUM formula, E9=SUM formula (0.10 points)
    try:
        c9_val = ws['C9'].value
        c9_bold = ws['C9'].font.bold
        d9_val = ws['D9'].value
        e9_val = ws['E9'].value

        total_checks = 0
        if c9_val is not None and str(c9_val).strip().lower() == 'total':
            total_checks += 1
        if c9_bold:
            total_checks += 1
        if d9_val is not None and isinstance(d9_val, str) and 'SUM' in d9_val.upper():
            total_checks += 1
        if e9_val is not None and isinstance(e9_val, str) and 'SUM' in e9_val.upper():
            total_checks += 1

        if total_checks == 4:
            print(f"PASS: Component 6 - Total row complete: C9='{c9_val}' bold={c9_bold}, D9='{d9_val}', E9='{e9_val}' (0.10 pts)")
            total_score += 0.10
        elif total_checks >= 2:
            partial = 0.05
            print(f"PARTIAL: Component 6 - {total_checks}/4 total row checks passed ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 6 - Total row: C9='{c9_val}' bold={c9_bold}, D9='{d9_val}', E9='{e9_val}'")
    except Exception as e:
        print(f"ERROR: Component 6 - {e}")

    # Component 7: Header formatting - dark green fill, white font on C1:F1 (0.05 points)
    try:
        format_ok = 0
        for coord in ['C1', 'D1', 'E1', 'F1']:
            cell = ws[coord]
            try:
                fill_rgb = cell.fill.fgColor.rgb if cell.fill.fgColor else None
            except Exception:
                fill_rgb = None
            try:
                font_rgb = cell.font.color.rgb if cell.font.color else None
            except Exception:
                font_rgb = None

            # Check for dark green fill (FF006100 or similar dark green)
            green_fill_detected = (
                fill_rgb is not None
                and isinstance(fill_rgb, str)
                and len(fill_rgb) >= 6
                and cell.fill.fill_type == 'solid'
            )
            if green_fill_detected:
                rgb_lower = fill_rgb.lower()
                r_hex = rgb_lower[2:4] if len(rgb_lower) >= 8 else rgb_lower[0:2]
                g_hex = rgb_lower[4:6] if len(rgb_lower) >= 8 else rgb_lower[2:4]
                b_hex = rgb_lower[6:8] if len(rgb_lower) >= 8 else rgb_lower[4:6]
                try:
                    r, g, b = int(r_hex, 16), int(g_hex, 16), int(b_hex, 16)
                    green_fill_detected = (g > r and g > b and r < 100)
                except ValueError:
                    green_fill_detected = False

            # Check white font
            white_font_detected = (
                font_rgb is not None
                and isinstance(font_rgb, str)
                and ('ffffff' in font_rgb.lower() or font_rgb.lower() in ('00ffffff', 'ffffffff'))
            )

            if green_fill_detected and white_font_detected:
                format_ok += 1

        if format_ok >= 3:
            print(f"PASS: Component 7 - {format_ok}/4 headers have dark green fill + white text (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 7 - Only {format_ok}/4 headers have correct fill+font color")
    except Exception as e:
        print(f"ERROR: Component 7 - {e}")

    # Component 8: Conditional formatting - data bars on D2:D7 and color scale on E2:E7 (0.10 points)
    try:
        data_bar_found = 0
        color_scale_found = 0
        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            for rule in cf.rules:
                if rule.type == 'dataBar' and ('D2' in cf_range or 'D' in cf_range):
                    data_bar_found += 1
                elif rule.type == 'colorScale' and ('E2' in cf_range or 'E' in cf_range):
                    color_scale_found += 1

        if data_bar_found >= 1 and color_scale_found >= 1:
            print(f"PASS: Component 8 - Data bars on D and color scale on E found (0.10 pts)")
            total_score += 0.10
        elif data_bar_found >= 1 or color_scale_found >= 1:
            partial = 0.05
            print(f"PARTIAL: Component 8 - dataBar={data_bar_found}, colorScale={color_scale_found} ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 8 - No conditional formatting found (dataBar={data_bar_found}, colorScale={color_scale_found})")
    except Exception as e:
        print(f"ERROR: Component 8 - {e}")

    # Component 9: Thick outside border around C1:F9 (0.05 points)
    try:
        border_ok = 0
        total_border_checks = 0

        # Top edge: C1:F1 should have thick top border
        for coord in ['C1', 'D1', 'E1', 'F1']:
            total_border_checks += 1
            if ws[coord].border.top.style in ('thick', 'medium'):
                border_ok += 1

        # Bottom edge: C9:F9 should have thick bottom border
        for coord in ['C9', 'D9', 'E9', 'F9']:
            total_border_checks += 1
            if ws[coord].border.bottom.style in ('thick', 'medium'):
                border_ok += 1

        # Left edge: C1:C9 should have thick left border
        for row in [1, 5, 9]:
            coord = f'C{row}'
            total_border_checks += 1
            if ws[coord].border.left.style in ('thick', 'medium'):
                border_ok += 1

        # Right edge: F1:F9 should have thick right border
        for row in [1, 5, 9]:
            coord = f'F{row}'
            total_border_checks += 1
            if ws[coord].border.right.style in ('thick', 'medium'):
                border_ok += 1

        if border_ok >= 10:
            print(f"PASS: Component 9 - Thick outside border detected ({border_ok}/{total_border_checks} checks) (0.05 pts)")
            total_score += 0.05
        elif border_ok >= 6:
            partial = 0.03
            print(f"PARTIAL: Component 9 - Partial border ({border_ok}/{total_border_checks} checks) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 9 - Thick outside border: {border_ok}/{total_border_checks} checks passed")
    except Exception as e:
        print(f"ERROR: Component 9 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
