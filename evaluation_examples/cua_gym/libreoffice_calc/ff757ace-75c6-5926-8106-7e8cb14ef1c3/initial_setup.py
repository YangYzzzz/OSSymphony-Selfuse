"""
Initial Setup: Create raw survey response time data for frequency distribution task.
Task ID: calc_gpm_024
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_024'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Survey ---
    ws = wb.active
    ws.title = 'Survey'

    # Header
    ws.cell(row=1, column=1, value='Response Time (s)')

    # 30 raw response time values
    values = [12, 15, 18, 22, 25, 28, 31, 34, 35, 38,
              41, 42, 45, 48, 50, 52, 55, 58, 60, 62,
              65, 68, 70, 72, 75, 80, 85, 90, 95, 120]

    for i, val in enumerate(values, 2):
        ws.cell(row=i, column=1, value=val)

    # Set column A width for readability
    ws.column_dimensions['A'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
