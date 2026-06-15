"""
Initial Setup: Set row 2 as repeating row for printing
Task ID: calc_gfl_052
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import random
from datetime import date, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_052'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # --- Row 1: Merged title ---
    ws.merge_cells("A1:F1")
    ws["A1"] = "Quarterly Sales Report 2025"
    ws["A1"].font = Font(name="Arial", size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # --- Row 2: Column headers ---
    headers = ["ID", "Name", "Date", "Amount", "Category", "Notes"]
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Column widths ---
    ws.column_dimensions["A"].width = 8    # ID
    ws.column_dimensions["B"].width = 22   # Name
    ws.column_dimensions["C"].width = 14   # Date
    ws.column_dimensions["D"].width = 14   # Amount
    ws.column_dimensions["E"].width = 16   # Category
    ws.column_dimensions["F"].width = 32   # Notes

    # --- Rows 3-120: 118 data entries ---
    first_names = [
        "Sarah", "Marcus", "Elena", "James", "Priya", "David", "Mei",
        "Carlos", "Aisha", "Thomas", "Yuki", "Robert", "Fatima", "Lucas",
        "Olga", "Ahmed", "Nina", "Patrick", "Lena", "Victor",
    ]
    last_names = [
        "Chen", "Johnson", "Petrov", "Williams", "Sharma", "Kim", "Garcia",
        "Okafor", "Mueller", "Santos", "Tanaka", "Brown", "Al-Rashid", "Johansson",
        "Kowalski", "Ibrahim", "Larsson", "O'Brien", "Fischer", "Moreau",
    ]
    categories = [
        "Electronics", "Office Supplies", "Software", "Furniture",
        "Travel", "Marketing", "Consulting", "Hardware", "Training", "Utilities",
    ]
    note_templates = [
        "Invoice #{inv} - {cat} purchase",
        "PO-{inv}: {cat} order placed",
        "Reimbursement for {cat} expense",
        "Vendor payment - {cat}",
        "Quarterly {cat} allocation",
        "Budget item #{inv}",
        "Approved by finance dept",
        "Recurring {cat} charge",
        "One-time {cat} procurement",
        "Project Alpha - {cat}",
    ]

    random.seed(42)
    base_date = date(2025, 1, 3)

    for i in range(118):
        row = i + 3
        entry_id = 1001 + i
        name = f"{random.choice(first_names)} {random.choice(last_names)}"
        entry_date = base_date + timedelta(days=random.randint(0, 270))
        amount = round(random.uniform(45.00, 12500.00), 2)
        category = random.choice(categories)
        note_tpl = random.choice(note_templates)
        note = note_tpl.format(inv=random.randint(10000, 99999), cat=category)

        ws.cell(row=row, column=1, value=entry_id)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=entry_date.strftime("%Y-%m-%d"))
        ws.cell(row=row, column=4, value=amount)
        ws.cell(row=row, column=4).number_format = '$#,##0.00'
        ws.cell(row=row, column=5, value=category)
        ws.cell(row=row, column=6, value=note)

    # NO print_title_rows set — that is the task for the agent

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
