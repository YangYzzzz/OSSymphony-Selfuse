"""
Reward Script: Set row 2 as repeating print row
Task ID: calc_gfl_052
Domain: libreoffice_calc
Scoring:
  Component 1 (0.6): print_title_rows is set to $2:$2
  Component 2 (0.4): print_title_rows is $2:$2 AND data integrity preserved
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_052'


def persist_app_state(domain: str):
    """Best-effort save in case LibreOffice has unsaved changes."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet 'Data' must exist
    if 'Data' not in wb.sheetnames:
        print("FAIL: Sheet 'Data' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Data']

    # Component 1: print_title_rows is set to $2:$2 (0.6 points)
    # This is the core task requirement: row 2 configured as repeating print row.
    try:
        ptr = ws.print_title_rows
        if ptr is not None:
            # Normalize: strip whitespace and compare
            ptr_normalized = str(ptr).strip().replace(' ', '')
            if ptr_normalized == '$2:$2':
                print(f"PASS: Component 1 — print_title_rows is '{ptr}' (exactly row 2) (0.6 pts)")
                total_score += 0.6
            else:
                print(f"FAIL: Component 1 — print_title_rows is '{ptr}', expected '$2:$2'")
        else:
            print("FAIL: Component 1 — print_title_rows is None (not set)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: print_title_rows is $2:$2 AND data integrity preserved (0.4 points)
    # This compound check ensures the task was done correctly without damaging the spreadsheet.
    # It only passes if the task change IS present (print_title_rows == $2:$2), so it
    # correctly returns 0 on the initial_env where print_title_rows is None.
    try:
        ptr = ws.print_title_rows
        ptr_ok = ptr is not None and str(ptr).strip().replace(' ', '') == '$2:$2'

        if not ptr_ok:
            print("FAIL: Component 2 — print_title_rows not set to $2:$2, skipping integrity check")
        else:
            integrity_issues = []

            # Check row count (should be 120 rows)
            if ws.max_row != 120:
                integrity_issues.append(f"max_row is {ws.max_row}, expected 120")

            # Check column count (should be 6 columns: A-F)
            if ws.max_column != 6:
                integrity_issues.append(f"max_column is {ws.max_column}, expected 6")

            # Check headers in row 2
            expected_headers = ['ID', 'Name', 'Date', 'Amount', 'Category', 'Notes']
            actual_headers = [ws.cell(row=2, column=c).value for c in range(1, 7)]
            if actual_headers != expected_headers:
                integrity_issues.append(f"headers are {actual_headers}, expected {expected_headers}")

            # Check merged cell A1:F1 exists
            from openpyxl.worksheet.merge import MergedCellRange
            merged_ranges = [str(r) for r in ws.merged_cells.ranges]
            if 'A1:F1' not in merged_ranges:
                integrity_issues.append(f"merged range A1:F1 not found, got {merged_ranges}")

            if len(integrity_issues) == 0:
                print(f"PASS: Component 2 — print_title_rows set AND data integrity preserved (0.4 pts)")
                total_score += 0.4
            else:
                print(f"FAIL: Component 2 — integrity issues: {'; '.join(integrity_issues)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
