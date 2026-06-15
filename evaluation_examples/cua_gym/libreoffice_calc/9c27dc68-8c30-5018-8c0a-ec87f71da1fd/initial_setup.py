"""
Initial Setup: Create a spreadsheet with product revenue data for pie chart task.
Task ID: calc_sales_087
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_087'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: ProductRev ---
    ws = wb.active
    ws.title = 'ProductRev'

    # Headers
    ws.cell(row=1, column=1, value='Product Line')
    ws.cell(row=1, column=2, value='Revenue')

    # Data rows
    data = [
        ['Enterprise SaaS', 520000],
        ['SMB SaaS', 280000],
        ['Professional Services', 175000],
        ['Hardware', 125000],
        ['Training', 60000],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 15

    # Format revenue as currency
    for r in range(2, 7):
        ws.cell(row=r, column=2).number_format = '#,##0'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
