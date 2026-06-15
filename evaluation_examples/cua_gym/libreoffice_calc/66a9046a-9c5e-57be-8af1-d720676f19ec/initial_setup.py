"""
Initial Setup: Commission clawback tracker with deal data
Task ID: calc_sales_084
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from datetime import date

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_084'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Clawback ---
    ws = wb.active
    ws.title = 'Clawback'

    # Headers
    headers = [
        'Deal', 'Close Date', 'Churn Date', 'Commission Paid',
        'Clawback Period (months)', 'Months Active', 'Months Remaining',
        'Clawback %', 'Clawback Amount'
    ]
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # Data rows
    data = [
        ['D1', date(2025, 1, 15), date(2025, 4, 15), 8000, 12],
        ['D2', date(2025, 3, 1),  date(2025, 5, 1),  5500, 12],
        ['D3', date(2025, 6, 1),  date(2025, 12, 1), 12000, 12],
        ['D4', date(2024, 8, 1),  date(2025, 9, 1),  9000, 12],
    ]

    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    date_fmt = 'yyyy-mm-dd'
    currency_fmt = '$#,##0'

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = data_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if c in (2, 3):  # date columns
                cell.number_format = date_fmt
            elif c == 4:  # commission
                cell.number_format = currency_fmt
            elif c == 5:  # period
                cell.number_format = '0'

    # Columns F-I are intentionally left empty (task is to add formulas)
    # Add borders to empty cells for visual consistency
    for r in range(2, 6):
        for c in range(6, 10):
            cell = ws.cell(row=r, column=c)
            cell.border = data_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set column widths for readability
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 18

    # Row height for header
    ws.row_dimensions[1].height = 30

    # Freeze header row
    ws.freeze_panes = 'A2'

    # --- Sheet: Policy ---
    ws2 = wb.create_sheet('Policy')
    ws2['A1'] = 'Commission Clawback Policy'
    ws2['A1'].font = Font(name='Calibri', size=14, bold=True)
    ws2['A3'] = 'Standard clawback period: 12 months from deal close date.'
    ws2['A4'] = 'If a deal churns within the clawback period, the commission is prorated.'
    ws2['A5'] = 'Formula: Clawback Amount = Commission Paid * (Months Remaining / Clawback Period)'
    ws2['A6'] = 'Months Active = months between Close Date and Churn Date.'
    ws2['A7'] = 'Months Remaining = MAX(Clawback Period - Months Active, 0).'
    ws2['A8'] = 'If Months Active >= Clawback Period, no clawback applies (Clawback Amount = $0).'
    ws2.column_dimensions['A'].width = 70

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
