"""
Reward Script: Commission clawback tracker with prorated refund calculations.
Task ID: calc_sales_084
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): F column — Months Active formulas present and logically correct
  Component 2 (0.25): G column — Months Remaining formulas present and logically correct
  Component 3 (0.25): H column — Clawback % formulas present and logically correct
  Component 4 (0.25): I column — Clawback Amount formulas present and logically correct
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_084'


def is_months_active_formula(val):
    """Check if value is a formula that calculates months between close and churn dates.
    Accepts DATEDIF-based or manual month calculation approaches."""
    if not isinstance(val, str) or not val.startswith('='):
        return False
    upper = val.upper().replace(" ", "")
    # DATEDIF approach: =DATEDIF(B,C,"M")
    if 'DATEDIF' in upper and '"M"' in upper:
        return True
    # Manual approaches: MONTH(C)-MONTH(B) or similar date arithmetic
    if 'MONTH' in upper and ('B' in upper or 'C' in upper):
        return True
    # YEARFRAC or DAYS-based approaches
    if 'YEARFRAC' in upper or ('DAYS' in upper and '30' in upper):
        return True
    # Integer division by 30 approaches
    if '(C' in upper and 'B' in upper and '30' in upper:
        return True
    return False


def is_months_remaining_formula(val, row):
    """Check if value is a formula for MAX(E-F, 0) or equivalent."""
    if not isinstance(val, str) or not val.startswith('='):
        return False
    upper = val.upper().replace(" ", "")
    r = str(row)
    # MAX(E-F, 0) pattern
    if 'MAX' in upper and ('E' + r) in upper and ('F' + r) in upper:
        return True
    # IF-based: =IF(E>F, E-F, 0)
    if 'IF' in upper and ('E' + r) in upper and ('F' + r) in upper:
        return True
    return False


def is_clawback_pct_formula(val, row):
    """Check if value is a formula for G/E (clawback percentage)."""
    if not isinstance(val, str) or not val.startswith('='):
        return False
    upper = val.upper().replace(" ", "")
    r = str(row)
    # G/E pattern (possibly wrapped in IF to avoid div-by-zero)
    if ('G' + r) in upper and ('E' + r) in upper:
        return True
    return False


def is_clawback_amount_formula(val, row):
    """Check if value is a formula for D*H (clawback amount)."""
    if not isinstance(val, str) or not val.startswith('='):
        return False
    upper = val.upper().replace(" ", "")
    r = str(row)
    # D*H pattern or ROUND(D*H,...) etc.
    if ('D' + r) in upper and ('H' + r) in upper:
        return True
    # Alternative: D * (G/E) directly
    if ('D' + r) in upper and ('G' + r) in upper and ('E' + r) in upper:
        return True
    return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Clawback' sheet must exist
    if 'Clawback' not in wb.sheetnames:
        print("FAIL: 'Clawback' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Clawback']

    # Component 1: F column — Months Active formulas (0.25 points)
    # Each row (2-5) with correct formula earns 0.0625 pts
    try:
        f_score = 0.0
        for row in range(2, 6):
            cell_val = ws.cell(row=row, column=6).value
            if is_months_active_formula(cell_val):
                print(f"  PASS: F{row} has months-active formula: {cell_val}")
                f_score += 0.0625
            else:
                print(f"  FAIL: F{row} expected months-active formula, found: {cell_val!r}")
        if f_score > 0:
            print(f"PASS: Component 1 — F column months active ({f_score:.4f} pts)")
            total_score += f_score
        else:
            print("FAIL: Component 1 — No valid months-active formulas in F2:F5")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: G column — Months Remaining formulas (0.25 points)
    try:
        g_score = 0.0
        for row in range(2, 6):
            cell_val = ws.cell(row=row, column=7).value
            if is_months_remaining_formula(cell_val, row):
                print(f"  PASS: G{row} has months-remaining formula: {cell_val}")
                g_score += 0.0625
            else:
                print(f"  FAIL: G{row} expected months-remaining formula, found: {cell_val!r}")
        if g_score > 0:
            print(f"PASS: Component 2 — G column months remaining ({g_score:.4f} pts)")
            total_score += g_score
        else:
            print("FAIL: Component 2 — No valid months-remaining formulas in G2:G5")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: H column — Clawback % formulas (0.25 points)
    try:
        h_score = 0.0
        for row in range(2, 6):
            cell_val = ws.cell(row=row, column=8).value
            if is_clawback_pct_formula(cell_val, row):
                print(f"  PASS: H{row} has clawback-% formula: {cell_val}")
                h_score += 0.0625
            else:
                print(f"  FAIL: H{row} expected clawback-% formula, found: {cell_val!r}")
        if h_score > 0:
            print(f"PASS: Component 3 — H column clawback % ({h_score:.4f} pts)")
            total_score += h_score
        else:
            print("FAIL: Component 3 — No valid clawback-% formulas in H2:H5")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: I column — Clawback Amount formulas (0.25 points)
    try:
        i_score = 0.0
        for row in range(2, 6):
            cell_val = ws.cell(row=row, column=9).value
            if is_clawback_amount_formula(cell_val, row):
                print(f"  PASS: I{row} has clawback-amount formula: {cell_val}")
                i_score += 0.0625
            else:
                print(f"  FAIL: I{row} expected clawback-amount formula, found: {cell_val!r}")
        if i_score > 0:
            print(f"PASS: Component 4 — I column clawback amount ({i_score:.4f} pts)")
            total_score += i_score
        else:
            print("FAIL: Component 4 — No valid clawback-amount formulas in I2:I5")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook for LibreOffice (save unsaved GUI edits)
def persist_app_state(domain):
    import os
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
