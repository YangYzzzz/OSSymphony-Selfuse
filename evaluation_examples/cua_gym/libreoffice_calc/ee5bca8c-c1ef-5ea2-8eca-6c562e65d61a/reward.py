"""
Reward Script: Calculate reorder quantity for items where current stock < reorder point
Task ID: calc_ops_003
Domain: libreoffice_calc
Scoring:
  Component 1 (0.2): All E2:E6 cells contain formulas (non-empty, start with =)
  Component 2 (0.5): Each formula has correct IF structure referencing correct cells
  Component 3 (0.3): Formulas use correct logic (IF stock < reorder_point, max - stock, 0)
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_003'


def normalize_formula(f):
    """Normalize formula for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ""
    return f.upper().replace(" ", "")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Reorder' sheet must exist
    if 'Reorder' not in wb.sheetnames:
        print(f"FAIL: Sheet 'Reorder' not found. Available: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Reorder']

    # Expected formulas for each row (row_num -> expected formula string)
    expected_formulas = {
        2: "=IF(B2<C2,D2-B2,0)",
        3: "=IF(B3<C3,D3-B3,0)",
        4: "=IF(B4<C4,D4-B4,0)",
        5: "=IF(B5<C5,D5-B5,0)",
        6: "=IF(B6<C6,D6-B6,0)",
    }

    # Expected computed values (for cross-check via formula logic)
    expected_values = {
        2: 350,   # 150 < 200 -> 500 - 150 = 350
        3: 0,     # 320 >= 250 -> 0
        4: 320,   # 80 < 100 -> 400 - 80 = 320
        5: 0,     # 500 >= 300 -> 0
        6: 255,   # 45 < 75 -> 300 - 45 = 255
    }

    # Component 1: All E2:E6 cells contain formulas (0.2 points)
    # This checks that formulas are present (not None/empty), which is the
    # fundamental change between initial (all None) and golden (all formulas).
    try:
        formula_count = 0
        for row in range(2, 7):
            val = ws.cell(row=row, column=5).value
            if val is not None and isinstance(val, str) and val.startswith('='):
                formula_count += 1
            else:
                print(f"  INFO: E{row} has no formula (value: {repr(val)})")

        if formula_count == 5:
            print(f"PASS: Component 1 — All 5 cells E2:E6 contain formulas (0.2 pts)")
            total_score += 0.2
        elif formula_count > 0:
            partial = round(0.2 * (formula_count / 5), 2)
            print(f"PARTIAL: Component 1 — {formula_count}/5 cells have formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No formulas found in E2:E6")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Each formula matches the expected IF structure (0.5 points, 0.1 per row)
    # Verifies the exact formula string for each cell.
    try:
        correct_formulas = 0
        for row in range(2, 7):
            val = ws.cell(row=row, column=5).value
            expected = expected_formulas[row]
            if val is not None and normalize_formula(val) == normalize_formula(expected):
                correct_formulas += 1
                print(f"  PASS: E{row} formula matches: {val}")
            else:
                print(f"  FAIL: E{row} expected '{expected}', found '{val}'")

        points = round(0.1 * correct_formulas, 2)
        if correct_formulas == 5:
            print(f"PASS: Component 2 — All 5 formulas match expected IF structure ({points} pts)")
            total_score += points
        elif correct_formulas > 0:
            print(f"PARTIAL: Component 2 — {correct_formulas}/5 formulas correct ({points} pts)")
            total_score += points
        else:
            print(f"FAIL: Component 2 — No formulas match expected structure")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Formulas encode correct reorder logic (0.3 points)
    # Cross-validates that formulas reference the correct cells to produce
    # the expected results. Checks that each formula contains an IF with
    # comparison of stock < reorder point and computes max_stock - current_stock.
    try:
        logic_correct = 0
        for row in range(2, 7):
            val = ws.cell(row=row, column=5).value
            if val is None or not isinstance(val, str):
                print(f"  FAIL: E{row} has no formula for logic check")
                continue

            norm = normalize_formula(val)
            # Check it's an IF formula with the right cell references
            b_ref = f"B{row}"
            c_ref = f"C{row}"
            d_ref = f"D{row}"

            has_if = "IF(" in norm
            has_comparison = f"{b_ref}<{c_ref}" in norm or f"{b_ref}<={c_ref}" in norm
            has_calc = f"{d_ref}-{b_ref}" in norm
            has_zero = norm.endswith(",0)") or ",0)" in norm

            if has_if and has_comparison and has_calc and has_zero:
                logic_correct += 1
                print(f"  PASS: E{row} has correct reorder logic")
            else:
                print(f"  FAIL: E{row} formula logic incorrect: IF={has_if}, comp={has_comparison}, calc={has_calc}, zero={has_zero}")

        points = round(0.3 * (logic_correct / 5), 2)
        if logic_correct == 5:
            print(f"PASS: Component 3 — All 5 formulas have correct reorder logic ({points} pts)")
            total_score += points
        elif logic_correct > 0:
            print(f"PARTIAL: Component 3 — {logic_correct}/5 have correct logic ({points} pts)")
            total_score += points
        else:
            print(f"FAIL: Component 3 — No formulas have correct reorder logic")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved LibreOffice state before verification
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
