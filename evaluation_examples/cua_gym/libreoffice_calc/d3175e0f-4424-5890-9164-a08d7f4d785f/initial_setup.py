"""
Initial Setup: Customer Loyalty Points System
Task ID: calc_wf_084
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_084'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

# --- Member data ---
MEMBER_NAMES = [
    "Sarah Chen", "Marcus Johnson", "Elena Rodriguez", "James Whitfield",
    "Aisha Patel", "David Kim", "Olivia Thompson", "Ryan Martinez",
    "Priya Sharma", "Lucas Weber", "Mei Lin Chang", "Carlos Gutierrez",
    "Hannah Brooks", "Tomasz Kowalski", "Fatima Al-Rashid", "Nathan Clarke",
    "Yuki Tanaka", "Rebecca Morrison", "Andre Williams", "Sophie Laurent"
]

MEMBER_IDS = [f"MEM-{1001 + i}" for i in range(20)]

random.seed(42)

# Generate join dates spread over 2023-2025
JOIN_DATES = []
base = datetime(2023, 1, 15)
for i in range(20):
    offset = random.randint(0, 700)
    JOIN_DATES.append(base + timedelta(days=offset))

# Assign tiers manually (initial state - these are the CURRENT tiers before task)
# These will be simple labels; the task asks the agent to create formulas
MEMBER_TIERS = [
    "Silver", "Gold", "Bronze", "Platinum", "Silver",
    "Gold", "Bronze", "Silver", "Gold", "Bronze",
    "Platinum", "Silver", "Bronze", "Gold", "Silver",
    "Bronze", "Gold", "Silver", "Bronze", "Silver"
]

# --- Transaction data (100 purchases) ---
# Distribute across members with varying amounts
TRANSACTIONS = []
random.seed(84)
tx_date_base = datetime(2023, 3, 1)
for i in range(100):
    member_idx = random.randint(0, 19)
    member_id = MEMBER_IDS[member_idx]
    tx_date = tx_date_base + timedelta(days=random.randint(0, 800))
    amount = round(random.uniform(15.0, 450.0), 2)
    points_earned = int(amount)  # 1 point per $1
    TRANSACTIONS.append((member_id, tx_date, amount, points_earned))

# Sort by date
TRANSACTIONS.sort(key=lambda x: x[1])

# --- Redemption data (25 entries) ---
REDEMPTIONS = []
random.seed(184)
redeem_base = datetime(2023, 6, 1)
redeemed_members = random.choices(range(20), k=25)
for i in range(25):
    member_idx = redeemed_members[i]
    member_id = MEMBER_IDS[member_idx]
    r_date = redeem_base + timedelta(days=random.randint(0, 700))
    points_used = random.randint(50, 500)
    REDEMPTIONS.append((member_id, r_date, points_used))

REDEMPTIONS.sort(key=lambda x: x[1])


def create_initial():
    wb = openpyxl.Workbook()

    # ==================== MEMBERS SHEET ====================
    ws_members = wb.active
    ws_members.title = "Members"

    headers = ["Member ID", "Name", "Join Date", "Current Tier"]
    for col, h in enumerate(headers, 1):
        cell = ws_members.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for r, (mid, name, jdate, tier) in enumerate(
        zip(MEMBER_IDS, MEMBER_NAMES, JOIN_DATES, MEMBER_TIERS), 2
    ):
        ws_members.cell(row=r, column=1, value=mid)
        ws_members.cell(row=r, column=2, value=name)
        ws_members.cell(row=r, column=3, value=jdate.strftime("%Y-%m-%d"))
        ws_members.cell(row=r, column=4, value=tier)

    ws_members.column_dimensions["A"].width = 14
    ws_members.column_dimensions["B"].width = 22
    ws_members.column_dimensions["C"].width = 14
    ws_members.column_dimensions["D"].width = 14

    # ==================== TRANSACTIONS SHEET ====================
    ws_tx = wb.create_sheet("Transactions")

    tx_headers = ["Member ID", "Date", "Amount", "Points Earned"]
    for col, h in enumerate(tx_headers, 1):
        cell = ws_tx.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for r, (mid, tdate, amount, pts) in enumerate(TRANSACTIONS, 2):
        ws_tx.cell(row=r, column=1, value=mid)
        ws_tx.cell(row=r, column=2, value=tdate.strftime("%Y-%m-%d"))
        c = ws_tx.cell(row=r, column=3, value=amount)
        c.number_format = '$#,##0.00'
        ws_tx.cell(row=r, column=4, value=pts)

    ws_tx.column_dimensions["A"].width = 14
    ws_tx.column_dimensions["B"].width = 14
    ws_tx.column_dimensions["C"].width = 14
    ws_tx.column_dimensions["D"].width = 16

    # ==================== REDEMPTIONS SHEET ====================
    ws_redeem = wb.create_sheet("Redemptions")

    redeem_headers = ["Member ID", "Date", "Points Used"]
    for col, h in enumerate(redeem_headers, 1):
        cell = ws_redeem.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for r, (mid, rdate, pts) in enumerate(REDEMPTIONS, 2):
        ws_redeem.cell(row=r, column=1, value=mid)
        ws_redeem.cell(row=r, column=2, value=rdate.strftime("%Y-%m-%d"))
        ws_redeem.cell(row=r, column=3, value=pts)

    ws_redeem.column_dimensions["A"].width = 14
    ws_redeem.column_dimensions["B"].width = 14
    ws_redeem.column_dimensions["C"].width = 14

    # ==================== DASHBOARD SHEET ====================
    ws_dash = wb.create_sheet("Dashboard")

    dash_headers = ["Member ID", "Name", "Current Balance", "Tier Status",
                    "Points to Next Tier"]
    for col, h in enumerate(dash_headers, 1):
        cell = ws_dash.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    ws_dash.column_dimensions["A"].width = 14
    ws_dash.column_dimensions["B"].width = 22
    ws_dash.column_dimensions["C"].width = 18
    ws_dash.column_dimensions["D"].width = 16
    ws_dash.column_dimensions["E"].width = 20

    # Dashboard is intentionally left with headers only - no data, no formulas,
    # no charts, no conditional formatting. The task asks the agent to populate it.

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
