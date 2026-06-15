"""
Reward Script: Customer loyalty points system
Task ID: calc_wf_084
Domain: libreoffice_calc
Scoring:
  Component 1: Dashboard populated with 20 member rows containing SUMIFS balance formulas (0.25)
  Component 2: Dashboard tier status column uses nested IF formulas (0.15)
  Component 3: Dashboard points-to-next-tier column uses IF formulas (0.10)
  Component 4: Conditional formatting rules for 4 tier colors on D2:D21 (0.20)
  Component 5: Bar chart present for member balances (0.15)
  Component 6: Pie chart present for tier distribution (0.15)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_084'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Dashboard sheet must exist
    if 'Dashboard' not in wb.sheetnames:
        print("FAIL: 'Dashboard' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Dashboard']

    # Component 1: Dashboard has 20 member data rows with SUMIFS balance formulas (0.25 points)
    # The initial Dashboard has only headers. Golden has rows 2-21 with SUMIFS formulas in column C.
    try:
        sumifs_count = 0
        data_row_count = 0
        for r in range(2, 22):
            member_id = ws.cell(row=r, column=1).value
            balance_val = ws.cell(row=r, column=3).value
            if member_id is not None and str(member_id).startswith('MEM-'):
                data_row_count += 1
                if balance_val is not None and isinstance(balance_val, str) and 'SUMIFS' in balance_val.upper():
                    sumifs_count += 1

        if data_row_count >= 18 and sumifs_count >= 18:
            print(f"PASS: Component 1 - {data_row_count} member rows with {sumifs_count} SUMIFS formulas (0.25 pts)")
            total_score += 0.25
        elif data_row_count >= 10 and sumifs_count >= 10:
            partial = 0.15
            print(f"PARTIAL: Component 1 - {data_row_count} member rows, {sumifs_count} SUMIFS formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - expected >=18 member rows with SUMIFS, found {data_row_count} rows, {sumifs_count} SUMIFS")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Tier status column (D) uses nested IF formulas with threshold logic (0.15 points)
    # Golden has formulas like =IF(C2>=5000,"Platinum",IF(C2>=2000,"Gold",IF(C2>=500,"Silver","Bronze")))
    try:
        tier_formula_count = 0
        for r in range(2, 22):
            tier_val = ws.cell(row=r, column=4).value
            if tier_val is not None and isinstance(tier_val, str):
                upper_val = tier_val.upper()
                if 'IF(' in upper_val and ('PLATINUM' in upper_val or 'GOLD' in upper_val or 'SILVER' in upper_val or 'BRONZE' in upper_val):
                    tier_formula_count += 1

        if tier_formula_count >= 18:
            print(f"PASS: Component 2 - {tier_formula_count} tier IF formulas found (0.15 pts)")
            total_score += 0.15
        elif tier_formula_count >= 10:
            partial = 0.08
            print(f"PARTIAL: Component 2 - {tier_formula_count} tier IF formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 - expected >=18 tier IF formulas, found {tier_formula_count}")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Points-to-next-tier column (E) uses IF formulas with threshold math (0.10 points)
    # Golden has formulas like =IF(C2>=5000,0,IF(C2>=2000,5000-C2,IF(C2>=500,2000-C2,500-C2)))
    try:
        next_tier_count = 0
        for r in range(2, 22):
            val = ws.cell(row=r, column=5).value
            if val is not None and isinstance(val, str):
                upper_val = val.upper()
                if 'IF(' in upper_val and ('5000' in val or '2000' in val or '500' in val):
                    next_tier_count += 1

        if next_tier_count >= 18:
            print(f"PASS: Component 3 - {next_tier_count} points-to-next-tier formulas found (0.10 pts)")
            total_score += 0.10
        elif next_tier_count >= 10:
            partial = 0.05
            print(f"PARTIAL: Component 3 - {next_tier_count} next-tier formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 - expected >=18 next-tier formulas, found {next_tier_count}")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Conditional formatting rules for tier colors on column D (0.20 points)
    # Golden has 4 rules: Bronze=brown, Silver=gray, Gold=yellow, Platinum=blue
    try:
        cf_rules = list(ws.conditional_formatting)
        tier_cf_count = 0
        tier_keywords_found = set()

        for cf in cf_rules:
            for rule in cf.rules:
                formula_str = str(getattr(rule, 'formula', '')).upper()
                if 'BRONZE' in formula_str:
                    tier_keywords_found.add('Bronze')
                    tier_cf_count += 1
                elif 'SILVER' in formula_str:
                    tier_keywords_found.add('Silver')
                    tier_cf_count += 1
                elif 'GOLD' in formula_str:
                    tier_keywords_found.add('Gold')
                    tier_cf_count += 1
                elif 'PLATINUM' in formula_str:
                    tier_keywords_found.add('Platinum')
                    tier_cf_count += 1

        if tier_cf_count >= 4 and len(tier_keywords_found) >= 4:
            print(f"PASS: Component 4 - {tier_cf_count} tier conditional formatting rules for {tier_keywords_found} (0.20 pts)")
            total_score += 0.20
        elif tier_cf_count >= 2:
            partial = 0.10
            print(f"PARTIAL: Component 4 - {tier_cf_count} rules for {tier_keywords_found} ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 - expected 4 tier CF rules, found {tier_cf_count} for {tier_keywords_found}")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Bar chart for member balances (0.15 points)
    # Golden has a BarChart with title containing "Balances" or "Loyalty"
    try:
        charts = ws._charts
        bar_chart_found = False
        for chart in charts:
            chart_class = chart.__class__.__name__
            if 'Bar' in chart_class:
                bar_chart_found = True
                break

        if bar_chart_found:
            print(f"PASS: Component 5 - Bar chart found on Dashboard (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 5 - No bar chart found on Dashboard (found {len(charts)} charts total)")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    # Component 6: Pie chart for tier distribution (0.15 points)
    # Golden has a PieChart with title containing "Tier Distribution"
    try:
        charts = ws._charts
        pie_chart_found = False
        for chart in charts:
            chart_class = chart.__class__.__name__
            if 'Pie' in chart_class:
                pie_chart_found = True
                break

        if pie_chart_found:
            print(f"PASS: Component 6 - Pie chart found on Dashboard (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 6 - No pie chart found on Dashboard (found {len(charts)} charts total)")
    except Exception as e:
        print(f"ERROR: Component 6 - {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state("libreoffice_calc")
    verify_task(file_path)
