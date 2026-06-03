"""
Initial Setup: Add subtotal rows, bold/border them, and format cost columns as currency
Task ID: calc_gsd_031
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_031'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    # Headers in row 1
    headers = ["SKU", "Product Name", "Category", "Qty in Stock",
               "Unit Cost", "Total Value", "Reorder Point"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Electronics (rows 2-11)
    electronics = [
        ["EL-1001", "Wireless Bluetooth Headphones", "Electronics", 145, 49.99, 7248.55, 30],
        ["EL-1002", "USB-C Charging Hub 7-Port", "Electronics", 89, 34.50, 3070.50, 20],
        ["EL-1003", "Mechanical Keyboard RGB", "Electronics", 67, 79.95, 5356.65, 15],
        ["EL-1004", "27-inch 4K Monitor", "Electronics", 23, 349.99, 8049.77, 10],
        ["EL-1005", "Portable SSD 1TB", "Electronics", 112, 89.99, 10078.88, 25],
        ["EL-1006", "Noise Cancelling Earbuds", "Electronics", 198, 129.00, 25542.00, 40],
        ["EL-1007", "Wireless Mouse Ergonomic", "Electronics", 234, 29.99, 7017.66, 50],
        ["EL-1008", "Laptop Stand Adjustable", "Electronics", 76, 44.50, 3382.00, 20],
        ["EL-1009", "Webcam 1080p HD", "Electronics", 158, 59.99, 9478.42, 35],
        ["EL-1010", "Smart Power Strip WiFi", "Electronics", 91, 24.99, 2274.09, 25],
    ]

    # Furniture (rows 12-21)
    furniture = [
        ["FN-2001", "Ergonomic Office Chair", "Furniture", 34, 289.00, 9826.00, 10],
        ["FN-2002", "Standing Desk Electric", "Furniture", 18, 549.99, 9899.82, 5],
        ["FN-2003", "Bookshelf 5-Tier Oak", "Furniture", 42, 129.50, 5439.00, 12],
        ["FN-2004", "Filing Cabinet 3-Drawer", "Furniture", 56, 89.99, 5039.44, 15],
        ["FN-2005", "Conference Table Oval", "Furniture", 8, 799.00, 6392.00, 3],
        ["FN-2006", "Desk Lamp LED Dimmable", "Furniture", 187, 39.99, 7478.13, 40],
        ["FN-2007", "Monitor Arm Dual Mount", "Furniture", 63, 74.50, 4693.50, 20],
        ["FN-2008", "Whiteboard Magnetic 4x6", "Furniture", 29, 119.99, 3479.71, 8],
        ["FN-2009", "Storage Ottoman Folding", "Furniture", 95, 54.99, 5224.05, 25],
        ["FN-2010", "Corner Desk L-Shaped", "Furniture", 14, 379.00, 5306.00, 5],
    ]

    # Clothing (rows 22-31)
    clothing = [
        ["CL-3001", "Cotton Polo Shirt Navy", "Clothing", 320, 24.99, 7996.80, 80],
        ["CL-3002", "Slim Fit Chinos Khaki", "Clothing", 185, 39.50, 7307.50, 50],
        ["CL-3003", "Merino Wool Sweater Grey", "Clothing", 78, 69.99, 5459.22, 20],
        ["CL-3004", "Rain Jacket Waterproof", "Clothing", 112, 89.00, 9968.00, 30],
        ["CL-3005", "Leather Belt Classic Brown", "Clothing", 245, 19.99, 4897.55, 60],
        ["CL-3006", "Running Shoes Lightweight", "Clothing", 156, 109.99, 17158.44, 40],
        ["CL-3007", "Denim Jacket Vintage Wash", "Clothing", 64, 74.50, 4768.00, 15],
        ["CL-3008", "Thermal Socks 3-Pack", "Clothing", 410, 12.99, 5325.90, 100],
        ["CL-3009", "Fleece Hoodie Zip-Up", "Clothing", 198, 44.99, 8908.02, 50],
        ["CL-3010", "Cargo Shorts Olive", "Clothing", 137, 29.99, 4108.63, 35],
    ]

    # Accessories (rows 32-41)
    accessories = [
        ["AC-4001", "Leather Wallet RFID Block", "Accessories", 289, 34.99, 10112.11, 70],
        ["AC-4002", "Stainless Steel Watch", "Accessories", 45, 149.99, 6749.55, 12],
        ["AC-4003", "Sunglasses Polarized UV400", "Accessories", 178, 29.99, 5338.22, 45],
        ["AC-4004", "Canvas Backpack 25L", "Accessories", 92, 59.99, 5519.08, 25],
        ["AC-4005", "Titanium Keychain Multi-tool", "Accessories", 334, 14.99, 5006.66, 80],
        ["AC-4006", "Silk Necktie Striped Blue", "Accessories", 67, 44.50, 2981.50, 18],
        ["AC-4007", "Travel Umbrella Compact", "Accessories", 215, 19.99, 4297.85, 55],
        ["AC-4008", "Phone Case Premium Leather", "Accessories", 423, 24.99, 10570.77, 100],
        ["AC-4009", "Beanie Hat Wool Blend", "Accessories", 156, 16.99, 2650.44, 40],
        ["AC-4010", "Crossbody Bag Nylon", "Accessories", 83, 39.99, 3319.17, 22],
    ]

    row = 2
    for dataset in [electronics, furniture, clothing, accessories]:
        for item in dataset:
            for col, val in enumerate(item, 1):
                ws.cell(row=row, column=col, value=val)
            row += 1

    # Set column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
