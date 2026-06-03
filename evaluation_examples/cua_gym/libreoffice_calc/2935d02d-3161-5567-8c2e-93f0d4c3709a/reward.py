"""
Reward Script: Inventory subtotal rows with formatting
Task ID: calc_gsd_031
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Four subtotal rows with correct category labels
  Component 2 (0.25): SUM formulas in Qty (D) and Total Value (F) columns
  Component 3 (0.25): Subtotal rows are bold with thin bottom border
  Component 4 (0.25): Currency formatting ($#,##0.00) on Unit Cost (E) and Total Value (F) data cells
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_031'

# Expected category subtotal labels
EXPECTED_SUBTOTALS = [
    'Electronics Subtotal',
    'Furniture Subtotal',
    'Clothing Subtotal',
    'Accessories Subtotal',
]


def find_subtotal_rows(ws):
    """Find rows whose column B contains a subtotal label. Returns dict: label -> row number."""
    found = {}
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=2).value
        if val and 'subtotal' in str(val).lower():
            found[str(val).strip()] = r
    return found


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Inventory' not in wb.sheetnames:
        print("FAIL: 'Inventory' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Inventory']

    # Find subtotal rows dynamically
    subtotal_map = find_subtotal_rows(ws)

    # Component 1: Four subtotal rows with correct category labels (0.25 points)
    try:
        matched = 0
        for label in EXPECTED_SUBTOTALS:
            # Allow case-insensitive and flexible matching
            found = False
            for actual_label, row in subtotal_map.items():
                if label.lower() == actual_label.lower():
                    found = True
                    break
                # Also accept label in column A
                a_val = ws.cell(row=row, column=1).value
                if a_val and label.lower() == str(a_val).strip().lower():
                    found = True
                    break
            if found:
                matched += 1
            else:
                print(f"FAIL: Component 1 -- Missing subtotal label: '{label}'")

        if matched == 4:
            print(f"PASS: Component 1 -- All 4 subtotal rows found with correct labels (0.25 pts)")
            total_score += 0.25
        elif matched > 0:
            partial = round(0.25 * matched / 4, 4)
            print(f"PARTIAL: Component 1 -- {matched}/4 subtotal labels found ({partial} pts)")
            total_score += partial
        else:
            print("FAIL: Component 1 -- No subtotal rows found")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: SUM formulas in D and F for each subtotal row (0.25 points)
    try:
        if not subtotal_map:
            print("FAIL: Component 2 -- No subtotal rows to check formulas")
        else:
            formula_checks = 0
            total_checks = 0
            for label in EXPECTED_SUBTOTALS:
                # Find the matching row
                row = None
                for actual_label, r in subtotal_map.items():
                    if label.lower() == actual_label.lower():
                        row = r
                        break
                if row is None:
                    continue

                for col in [4, 6]:  # D=Qty, F=Total Value
                    total_checks += 1
                    cell_val = ws.cell(row=row, column=col).value
                    if cell_val and isinstance(cell_val, str) and '=SUM(' in cell_val.upper().replace(' ', ''):
                        formula_checks += 1
                    else:
                        col_letter = 'D' if col == 4 else 'F'
                        print(f"FAIL: Component 2 -- Row {row} ({label}): {col_letter}{row} = {cell_val!r}, expected SUM formula")

            if total_checks > 0 and formula_checks == total_checks:
                print(f"PASS: Component 2 -- All {formula_checks} SUM formulas present (0.25 pts)")
                total_score += 0.25
            elif formula_checks > 0:
                partial = round(0.25 * formula_checks / max(total_checks, 1), 4)
                print(f"PARTIAL: Component 2 -- {formula_checks}/{total_checks} SUM formulas found ({partial} pts)")
                total_score += partial
            else:
                print("FAIL: Component 2 -- No SUM formulas found in subtotal rows")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Subtotal rows are bold with bottom border (0.25 points)
    try:
        if not subtotal_map:
            print("FAIL: Component 3 -- No subtotal rows to check formatting")
        else:
            style_pass = 0
            style_total = 0
            for label in EXPECTED_SUBTOTALS:
                row = None
                for actual_label, r in subtotal_map.items():
                    if label.lower() == actual_label.lower():
                        row = r
                        break
                if row is None:
                    continue

                style_total += 1
                # Check bold on column B (the label cell)
                label_cell = ws.cell(row=row, column=2)
                is_bold = label_cell.font.bold is True

                # Check bottom border on any cell in the row
                has_bottom_border = False
                for c in range(1, 8):
                    border = ws.cell(row=row, column=c).border
                    if border.bottom and border.bottom.style is not None:
                        has_bottom_border = True
                        break

                if is_bold and has_bottom_border:
                    style_pass += 1
                else:
                    reasons = []
                    if not is_bold:
                        reasons.append("not bold")
                    if not has_bottom_border:
                        reasons.append("no bottom border")
                    print(f"FAIL: Component 3 -- Row {row} ({label}): {', '.join(reasons)}")

            if style_total > 0 and style_pass == style_total:
                print(f"PASS: Component 3 -- All {style_pass} subtotal rows are bold with bottom border (0.25 pts)")
                total_score += 0.25
            elif style_pass > 0:
                partial = round(0.25 * style_pass / max(style_total, 1), 4)
                print(f"PARTIAL: Component 3 -- {style_pass}/{style_total} subtotal rows correctly styled ({partial} pts)")
                total_score += partial
            else:
                print("FAIL: Component 3 -- No subtotal rows have correct formatting")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Currency formatting on E and F data cells (0.25 points)
    # Check that Unit Cost (E) and Total Value (F) data cells are formatted as USD currency
    # We check a sample of data rows (non-header, non-subtotal)
    try:
        subtotal_row_set = set(subtotal_map.values())
        currency_ok = 0
        currency_total = 0

        # Check E and F columns for all data rows (skip header row 1 and subtotal rows)
        for r in range(2, ws.max_row + 1):
            if r in subtotal_row_set:
                # For subtotal rows, only F needs currency (E is empty in golden)
                # But the task says "format cost column as currency throughout"
                # so check F on subtotal rows too
                cell_f = ws.cell(row=r, column=6)
                if cell_f.value is not None:
                    currency_total += 1
                    nf = cell_f.number_format or ''
                    if '$' in nf:
                        currency_ok += 1
                continue

            # Regular data rows: check both E and F
            for col in [5, 6]:
                cell = ws.cell(row=r, column=col)
                if cell.value is not None:
                    currency_total += 1
                    nf = cell.number_format or ''
                    if '$' in nf:
                        currency_ok += 1

        if currency_total > 0:
            ratio = currency_ok / currency_total
            if ratio >= 0.9:
                print(f"PASS: Component 4 -- {currency_ok}/{currency_total} cells have USD currency format (0.25 pts)")
                total_score += 0.25
            elif ratio >= 0.5:
                partial = round(0.25 * ratio, 4)
                print(f"PARTIAL: Component 4 -- {currency_ok}/{currency_total} cells formatted as currency ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 4 -- Only {currency_ok}/{currency_total} cells have currency format")
        else:
            print("FAIL: Component 4 -- No data cells found to check currency format")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
