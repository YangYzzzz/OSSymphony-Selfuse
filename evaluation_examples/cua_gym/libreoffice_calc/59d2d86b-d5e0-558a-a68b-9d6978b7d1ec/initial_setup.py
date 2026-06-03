"""
Initial Setup: HR Exit Interview Tracker
Task ID: calc_hr_exit_interview_tracker_052
Domain: libreoffice_calc

Creates initial file with:
- Sheet 'Exit Interviews': 55 rows of exit interview data with inconsistent free-text in E/F
- Sheet 'Exit Summary': empty
"""

import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_exit_interview_tracker_052'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Exit Interviews ---
    ws1 = wb.active
    ws1.title = 'Exit Interviews'

    # Headers (row 1)
    headers = ['Emp ID', 'Name', 'Department', 'Last Day', 'Departure Reason', 'Regrettable']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # Realistic exit interview data (rows 2-56 = 55 records)
    # Columns: Emp ID, Name, Department, Last Day, Departure Reason (free-text, inconsistent), Regrettable (free-text)
    data = [
        ['EMP-1001', 'Sarah Chen', 'Engineering', '2024-01-15', 'Better pay elsewhere', 'yes'],
        ['EMP-1002', 'Marcus Johnson', 'Marketing', '2024-01-22', 'Career Growth', 'Yes'],
        ['EMP-1003', 'Priya Patel', 'HR', '2024-01-31', 'relocation', 'No'],
        ['EMP-1004', 'Derek Williams', 'Finance', '2024-02-05', 'compensation issues', 'YES'],
        ['EMP-1005', 'Amanda Torres', 'Engineering', '2024-02-12', 'Management', 'Yes'],
        ['EMP-1006', 'James Lee', 'Sales', '2024-02-19', 'culture fit', 'no'],
        ['EMP-1007', 'Nina Sharma', 'Engineering', '2024-02-28', 'Career advancement', 'Yes'],
        ['EMP-1008', 'Robert Kim', 'Operations', '2024-03-04', 'moving cities', 'No'],
        ['EMP-1009', 'Lisa Nguyen', 'Marketing', '2024-03-11', 'Comp', 'YES'],
        ['EMP-1010', 'Carlos Mendez', 'Finance', '2024-03-18', 'poor management', 'Yes'],
        ['EMP-1011', 'Emily Watson', 'HR', '2024-03-25', 'other', 'no'],
        ['EMP-1012', 'David Park', 'Engineering', '2024-04-01', 'salary', 'Yes'],
        ['EMP-1013', 'Michelle Brown', 'Sales', '2024-04-08', 'Career Growth', 'Yes'],
        ['EMP-1014', 'Kevin Okafor', 'Operations', '2024-04-15', 'company culture', 'No'],
        ['EMP-1015', 'Rachel Green', 'Engineering', '2024-04-22', 'better opportunity', 'Yes'],
        ['EMP-1016', 'Thomas Wright', 'Marketing', '2024-04-29', 'relocation', 'No'],
        ['EMP-1017', 'Sandra Liu', 'Finance', '2024-05-06', 'Management issues', 'Yes'],
        ['EMP-1018', 'Brian Davis', 'Engineering', '2024-05-13', 'COMPENSATION', 'Yes'],
        ['EMP-1019', 'Jennifer Adams', 'HR', '2024-05-20', 'personal reasons', 'No'],
        ['EMP-1020', 'Michael Scott', 'Sales', '2024-05-27', 'culture', 'no'],
        ['EMP-1021', 'Patricia Hall', 'Operations', '2024-06-03', 'career growth', 'YES'],
        ['EMP-1022', 'Christopher Jones', 'Engineering', '2024-06-10', 'Mgmt', 'Yes'],
        ['EMP-1023', 'Angela White', 'Marketing', '2024-06-17', 'other opportunities', 'No'],
        ['EMP-1024', 'Steven Clark', 'Finance', '2024-06-24', 'pay', 'Yes'],
        ['EMP-1025', 'Diana Lewis', 'HR', '2024-07-01', 'Relocation', 'No'],
        ['EMP-1026', 'Jason Robinson', 'Engineering', '2024-07-08', 'growth opportunities', 'Yes'],
        ['EMP-1027', 'Karen Walker', 'Sales', '2024-07-15', 'Culture', 'no'],
        ['EMP-1028', 'Andrew Hill', 'Operations', '2024-07-22', 'management style', 'Yes'],
        ['EMP-1029', 'Laura Martinez', 'Engineering', '2024-07-29', 'Compensation', 'YES'],
        ['EMP-1030', 'Paul Thompson', 'Marketing', '2024-08-05', 'career change', 'No'],
        ['EMP-1031', 'Megan Garcia', 'Finance', '2024-08-12', 'other', 'No'],
        ['EMP-1032', 'Daniel Wilson', 'HR', '2024-08-19', 'better salary', 'Yes'],
        ['EMP-1033', 'Stephanie Moore', 'Engineering', '2024-08-26', 'career path', 'Yes'],
        ['EMP-1034', 'Eric Taylor', 'Sales', '2024-09-02', 'relocation', 'No'],
        ['EMP-1035', 'Rebecca Anderson', 'Operations', '2024-09-09', 'management', 'Yes'],
        ['EMP-1036', 'Jonathan Jackson', 'Engineering', '2024-09-16', 'Culture issues', 'no'],
        ['EMP-1037', 'Heather Harris', 'Marketing', '2024-09-23', 'CAREER GROWTH', 'Yes'],
        ['EMP-1038', 'Ryan Martinez', 'Finance', '2024-09-30', 'Compensation', 'Yes'],
        ['EMP-1039', 'Amy Nelson', 'HR', '2024-10-07', 'moving', 'No'],
        ['EMP-1040', 'Brandon Carter', 'Engineering', '2024-10-14', 'better offer', 'Yes'],
        ['EMP-1041', 'Melissa Mitchell', 'Sales', '2024-10-21', 'company culture', 'no'],
        ['EMP-1042', 'Tyler Perez', 'Operations', '2024-10-28', 'Management', 'Yes'],
        ['EMP-1043', 'Samantha Roberts', 'Engineering', '2024-11-04', 'salary increase', 'YES'],
        ['EMP-1044', 'Jacob Turner', 'Marketing', '2024-11-11', 'other', 'No'],
        ['EMP-1045', 'Kayla Phillips', 'Finance', '2024-11-18', 'career advancement', 'Yes'],
        ['EMP-1046', 'Nathan Campbell', 'HR', '2024-11-25', 'Culture', 'No'],
        ['EMP-1047', 'Ashley Parker', 'Engineering', '2024-12-02', 'compensation', 'Yes'],
        ['EMP-1048', 'Zachary Evans', 'Sales', '2024-12-09', 'Relocation', 'No'],
        ['EMP-1049', 'Jessica Collins', 'Operations', '2024-12-16', 'career growth', 'Yes'],
        ['EMP-1050', 'Austin Edwards', 'Engineering', '2024-12-23', 'management problems', 'Yes'],
        ['EMP-1051', 'Madison Stewart', 'Marketing', '2025-01-06', 'other', 'no'],
        ['EMP-1052', 'Ethan Morris', 'Finance', '2025-01-13', 'pay raise needed', 'Yes'],
        ['EMP-1053', 'Brittany Rogers', 'HR', '2025-01-20', 'relocation', 'No'],
        ['EMP-1054', 'Caleb Reed', 'Engineering', '2025-01-27', 'Career Growth', 'YES'],
        ['EMP-1055', 'Hannah Cook', 'Sales', '2025-02-03', 'culture mismatch', 'No'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Column widths for readability
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 16
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 24
    ws1.column_dimensions['F'].width = 14

    # --- Sheet 2: Exit Summary (empty) ---
    ws2 = wb.create_sheet('Exit Summary')
    # Intentionally empty — task asks agent to populate it

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
