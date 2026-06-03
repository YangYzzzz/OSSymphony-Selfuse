"""
Reward Script: Exit Interview Tracker Setup
Task ID: calc_hr_exit_interview_tracker_052
Domain: libreoffice_calc
Scoring:
  Component 1: Departure Reason dropdown (E2:E56) — 0.25 pts
  Component 2: Regrettable dropdown (F2:F56) — 0.25 pts
  Component 3: Exit Summary headers (A1:D1, bold) — 0.20 pts
  Component 4: Exit Summary department rows with correct formulas — 0.20 pts
  Component 5: Exit Summary totals row (bold) + D column % format — 0.10 pts
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_exit_interview_tracker_052'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # -----------------------------------------------------------------------
    # Component 1: Departure Reason dropdown validation on E2:E56 (0.25 pts)
    # The initial file has NO data validations. Golden adds dropdown for E col.
    # Expected formula1: "Compensation,Career Growth,Management,Culture,Relocation,Other"
    # -----------------------------------------------------------------------
    try:
        if 'Exit Interviews' not in wb.sheetnames:
            print("FAIL: Component 1 — 'Exit Interviews' sheet not found")
        else:
            ws_ei = wb['Exit Interviews']
            dvs = ws_ei.data_validations.dataValidation

            # Look for a list validation covering column E
            expected_options = {"Compensation", "Career Growth", "Management", "Culture", "Relocation", "Other"}
            found_e_dv = False

            for dv in dvs:
                if dv.type != 'list':
                    continue
                sqref_str = str(dv.sqref)
                # Check if it covers E column cells
                if 'E' not in sqref_str:
                    continue
                # Check formula1 contains the correct options
                formula = dv.formula1.strip('"').strip("'")
                actual_options = {o.strip() for o in formula.split(',')}
                if actual_options == expected_options:
                    found_e_dv = True
                    print(f"PASS: Component 1 — Departure Reason dropdown found on {sqref_str} (0.25 pts)")
                    total_score += 0.25
                    break
                else:
                    print(f"FAIL: Component 1 — Dropdown options mismatch. Expected: {expected_options}, Found: {actual_options}")
                    break

            if not found_e_dv and total_score < 0.25:
                # Check if we printed a message already
                has_e_dv = any(dv.type == 'list' and 'E' in str(dv.sqref) for dv in dvs)
                if not has_e_dv:
                    print(f"FAIL: Component 1 — No list validation found on column E")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Regrettable dropdown validation on F2:F56 (0.25 pts)
    # Expected formula1: "Yes,No"
    # -----------------------------------------------------------------------
    try:
        if 'Exit Interviews' not in wb.sheetnames:
            print("FAIL: Component 2 — 'Exit Interviews' sheet not found")
        else:
            ws_ei = wb['Exit Interviews']
            dvs = ws_ei.data_validations.dataValidation

            expected_options_f = {"Yes", "No"}
            found_f_dv = False

            for dv in dvs:
                if dv.type != 'list':
                    continue
                sqref_str = str(dv.sqref)
                if 'F' not in sqref_str:
                    continue
                formula = dv.formula1.strip('"').strip("'")
                actual_options = {o.strip() for o in formula.split(',')}
                if actual_options == expected_options_f:
                    found_f_dv = True
                    print(f"PASS: Component 2 — Regrettable dropdown found on {sqref_str} (0.25 pts)")
                    total_score += 0.25
                    break
                else:
                    print(f"FAIL: Component 2 — Dropdown options mismatch. Expected: {expected_options_f}, Found: {actual_options}")
                    break

            if not found_f_dv and total_score < 0.50:
                has_f_dv = any(dv.type == 'list' and 'F' in str(dv.sqref) for dv in dvs)
                if not has_f_dv:
                    print(f"FAIL: Component 2 — No list validation found on column F")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Exit Summary headers (A1:D1, bold) (0.20 pts)
    # Initial has empty Exit Summary. Golden adds bold headers.
    # -----------------------------------------------------------------------
    try:
        if 'Exit Summary' not in wb.sheetnames:
            print("FAIL: Component 3 — 'Exit Summary' sheet not found")
        else:
            ws_es = wb['Exit Summary']
            expected_headers = {1: 'Department', 2: 'Total Exits', 3: 'Regrettable', 4: '% Regrettable'}

            headers_ok = True
            for col, expected in expected_headers.items():
                cell = ws_es.cell(row=1, column=col)
                actual = cell.value
                if actual != expected:
                    print(f"FAIL: Component 3 — Header col {col}: expected '{expected}', found '{actual}'")
                    headers_ok = False
                    break
                if not cell.font.bold:
                    print(f"FAIL: Component 3 — Header col {col} ('{expected}') is not bold")
                    headers_ok = False
                    break

            if headers_ok:
                print(f"PASS: Component 3 — Exit Summary headers A1:D1 present and bold (0.20 pts)")
                total_score += 0.20
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: Exit Summary department rows 2-7 with correct formulas (0.20 pts)
    # Each row should have: dept name in A, COUNTIF formula in B, COUNTIFS in C, IFERROR in D
    # -----------------------------------------------------------------------
    try:
        if 'Exit Summary' not in wb.sheetnames:
            print("FAIL: Component 4 — 'Exit Summary' sheet not found")
        else:
            ws_es = wb['Exit Summary']
            expected_depts = ['Engineering', 'Marketing', 'HR', 'Finance', 'Sales', 'Operations']

            # Check that rows 2-7 exist and have formulas in B, C, D columns
            rows_ok = True
            dept_rows_found = 0

            for row in range(2, 8):
                dept_cell = ws_es.cell(row=row, column=1).value
                b_cell = ws_es.cell(row=row, column=2).value
                c_cell = ws_es.cell(row=row, column=3).value
                d_cell = ws_es.cell(row=row, column=4).value

                # Department name check (must be a non-empty string that is a known dept)
                if not dept_cell or not isinstance(dept_cell, str):
                    print(f"FAIL: Component 4 — Row {row}: no department name in A{row}")
                    rows_ok = False
                    break

                # B column: must contain COUNTIF formula referencing Exit Interviews.$C:$C
                if not b_cell or not isinstance(b_cell, str):
                    print(f"FAIL: Component 4 — Row {row}: B{row} has no formula, found: {repr(b_cell)}")
                    rows_ok = False
                    break
                b_upper = b_cell.upper().replace(' ', '')
                if 'COUNTIF' not in b_upper:
                    print(f"FAIL: Component 4 — Row {row}: B{row} should have COUNTIF formula, found: {repr(b_cell)}")
                    rows_ok = False
                    break

                # C column: must contain COUNTIFS formula
                if not c_cell or not isinstance(c_cell, str):
                    print(f"FAIL: Component 4 — Row {row}: C{row} has no formula, found: {repr(c_cell)}")
                    rows_ok = False
                    break
                c_upper = c_cell.upper().replace(' ', '')
                if 'COUNTIFS' not in c_upper:
                    print(f"FAIL: Component 4 — Row {row}: C{row} should have COUNTIFS formula, found: {repr(c_cell)}")
                    rows_ok = False
                    break

                # D column: must contain IFERROR formula
                if not d_cell or not isinstance(d_cell, str):
                    print(f"FAIL: Component 4 — Row {row}: D{row} has no formula, found: {repr(d_cell)}")
                    rows_ok = False
                    break
                d_upper = d_cell.upper().replace(' ', '')
                if 'IFERROR' not in d_upper:
                    print(f"FAIL: Component 4 — Row {row}: D{row} should have IFERROR formula, found: {repr(d_cell)}")
                    rows_ok = False
                    break

                dept_rows_found += 1

            if rows_ok and dept_rows_found == 6:
                print(f"PASS: Component 4 — All 6 department rows (2-7) have correct COUNTIF/COUNTIFS/IFERROR formulas (0.20 pts)")
                total_score += 0.20
            elif dept_rows_found > 0:
                print(f"PARTIAL: Component 4 — Only {dept_rows_found}/6 department rows have correct formulas")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -----------------------------------------------------------------------
    # Component 5: Totals row (row 8) is bold AND D column formatted as percentage (0.10 pts)
    # -----------------------------------------------------------------------
    try:
        if 'Exit Summary' not in wb.sheetnames:
            print("FAIL: Component 5 — 'Exit Summary' sheet not found")
        else:
            ws_es = wb['Exit Summary']

            # Check totals row (row 8) is bold
            totals_bold = True
            for col in range(1, 5):
                cell = ws_es.cell(row=8, column=col)
                if cell.value is not None and not cell.font.bold:
                    totals_bold = False
                    print(f"FAIL: Component 5 — Totals row cell {cell.coordinate} is not bold")
                    break

            # Also check row 8 has a dept label 'Total' and SUM formulas in B8, C8
            a8 = ws_es.cell(row=8, column=1).value
            b8 = ws_es.cell(row=8, column=2).value
            has_totals = (a8 is not None and
                          b8 is not None and isinstance(b8, str) and 'SUM' in b8.upper())

            # Check D2:D7 are percentage formatted
            pct_formats = {'0%', '0.00%', '0.0%', '#,##0.00%', '#,##0%'}
            d_pct_ok = True
            for row in range(2, 9):
                cell = ws_es.cell(row=row, column=4)
                if cell.value is not None and isinstance(cell.value, str) and 'IFERROR' in cell.value.upper():
                    # This is a formula cell, check its number format
                    nf = cell.number_format
                    if not any(pct in nf for pct in ['%']):
                        d_pct_ok = False
                        print(f"FAIL: Component 5 — D{row} not percentage formatted, number_format='{nf}'")
                        break

            if totals_bold and has_totals and d_pct_ok:
                print(f"PASS: Component 5 — Totals row (row 8) is bold + D column has % format (0.10 pts)")
                total_score += 0.10
            else:
                if not totals_bold:
                    print(f"FAIL: Component 5 — Totals row is not fully bold")
                if not has_totals:
                    print(f"FAIL: Component 5 — Totals row A8='{a8}', B8={repr(b8)} — expected 'Total' label and SUM formula")
                if not d_pct_ok:
                    print(f"FAIL: Component 5 — D column not formatted as percentage")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
