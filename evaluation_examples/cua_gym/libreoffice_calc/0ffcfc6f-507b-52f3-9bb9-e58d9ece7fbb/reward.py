"""
Reward Script: Lead Scoring Worksheet with Weighted Formula and Grade Assignment
Task ID: calc_sales_068
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): G2:G5 contain scoring formulas with correct weighted criteria
  Component 2 (0.3): G column formulas use correct weight values for all 5 criteria
  Component 3 (0.3): H2:H5 contain grade assignment formulas based on G column
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_068'

# Expected ground truth values from task context
EXPECTED_TOTALS = {
    2: 97.75,   # Lead 1
    3: 14.5,    # Lead 2
    4: 74.75,   # Lead 3
    5: 80.8,    # Lead 4
}
EXPECTED_GRADES = {
    2: 'A',
    3: 'D',
    4: 'B',
    5: 'A',
}

# Scoring weights from task context
SIZE_WEIGHTS = {'Enterprise': 30, 'Mid-Market': 20, 'SMB': 10}
INDUSTRY_WEIGHT = 20   # Yes=20, No=0
BUDGET_WEIGHT = 15     # Yes=15, No=0
TIMELINE_WEIGHTS = {'<3 months': 20, '3-6 months': 10, '>6 months': 0}
ENGAGEMENT_FACTOR = 15  # F/100*15


def normalize_formula(f):
    """Normalize a formula string for comparison: uppercase, remove spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def formula_has_scoring_structure(formula_str, row):
    """
    Check if a G-column formula contains the key scoring components:
    - Company size IF logic (Enterprise/Mid-Market/SMB)
    - Industry match check
    - Budget check
    - Timeline check
    - Engagement calculation (F/100*15 or equivalent)
    """
    f = normalize_formula(formula_str)
    if not f.startswith('='):
        return False, "Does not start with ="

    checks_passed = 0
    details = []

    # Check for company size scoring (Enterprise, Mid-Market, SMB)
    if 'ENTERPRISE' in f and 'MID-MARKET' in f and 'SMB' in f:
        checks_passed += 1
        details.append("size scoring present")
    else:
        details.append("MISSING size scoring (Enterprise/Mid-Market/SMB)")

    # Check for industry match
    if f'C{row}' in f and 'YES' in f:
        checks_passed += 1
        details.append("industry match present")
    else:
        details.append(f"MISSING industry match (C{row})")

    # Check for budget confirmed
    if f'D{row}' in f:
        checks_passed += 1
        details.append("budget check present")
    else:
        details.append(f"MISSING budget check (D{row})")

    # Check for timeline
    if f'E{row}' in f and ('<3' in formula_str or '3-6' in formula_str or '<3MONTHS' in f):
        checks_passed += 1
        details.append("timeline check present")
    else:
        details.append(f"MISSING timeline check (E{row})")

    # Check for engagement (F/100*15 or similar)
    if f'F{row}' in f and ('100' in f) and ('15' in f):
        checks_passed += 1
        details.append("engagement calc present")
    else:
        details.append(f"MISSING engagement calc (F{row}/100*15)")

    return checks_passed, details


def formula_has_correct_weights(formula_str):
    """
    Check if the formula contains the correct weight values:
    30 (Enterprise), 20 (Mid-Market/Industry/Timeline<3), 10 (SMB/Timeline3-6), 15 (Budget/Engagement)
    """
    f = normalize_formula(formula_str)
    checks = 0
    details = []

    # Check for weight 30 (Enterprise)
    if '30' in f:
        checks += 1
        details.append("30 (Enterprise) found")
    else:
        details.append("MISSING 30 (Enterprise)")

    # Check for weight 20 (Mid-Market and Industry and Timeline<3)
    if '20' in f:
        checks += 1
        details.append("20 (Mid-Market/Industry/Timeline) found")
    else:
        details.append("MISSING 20")

    # Check for weight 10 (SMB and Timeline 3-6)
    if '10' in f:
        checks += 1
        details.append("10 (SMB/Timeline3-6) found")
    else:
        details.append("MISSING 10")

    # Check for weight 15 (Budget and Engagement)
    if '15' in f:
        checks += 1
        details.append("15 (Budget/Engagement) found")
    else:
        details.append("MISSING 15")

    return checks, details


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check LeadScoring sheet exists
    if 'LeadScoring' not in wb.sheetnames:
        print("CRITICAL: 'LeadScoring' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['LeadScoring']

    # Component 1: G2:G5 contain scoring formulas with correct structure (0.4 points)
    # Each row contributes 0.1 points
    try:
        comp1_score = 0.0
        for row in range(2, 6):
            cell_val = ws.cell(row=row, column=7).value  # G column
            if cell_val is None:
                print(f"FAIL: G{row} is empty — no scoring formula")
                continue
            if not isinstance(cell_val, str) or not cell_val.startswith('='):
                print(f"FAIL: G{row} is not a formula: {cell_val}")
                continue

            checks, details = formula_has_scoring_structure(cell_val, row)
            if checks >= 4:  # At least 4 of 5 scoring components present
                comp1_score += 0.1
                print(f"PASS: G{row} has scoring formula ({checks}/5 components): {', '.join(details)}")
            else:
                print(f"FAIL: G{row} scoring formula incomplete ({checks}/5 components): {', '.join(details)}")

        if comp1_score > 0:
            print(f"PASS: Component 1 — G column scoring formulas ({comp1_score} pts)")
            total_score += comp1_score
        else:
            print(f"FAIL: Component 1 — No valid scoring formulas in G column")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: G column formulas use correct weight values (0.3 points)
    # Each row contributes 0.075 points
    try:
        comp2_score = 0.0
        for row in range(2, 6):
            cell_val = ws.cell(row=row, column=7).value  # G column
            if cell_val is None or not isinstance(cell_val, str):
                print(f"FAIL: G{row} no formula to check weights")
                continue

            checks, details = formula_has_correct_weights(cell_val)
            if checks >= 3:  # At least 3 of 4 weight groups correct
                comp2_score += 0.075
                print(f"PASS: G{row} has correct weights ({checks}/4): {', '.join(details)}")
            else:
                print(f"FAIL: G{row} incorrect weights ({checks}/4): {', '.join(details)}")

        if comp2_score > 0:
            print(f"PASS: Component 2 — G column correct weights ({comp2_score} pts)")
            total_score += comp2_score
        else:
            print(f"FAIL: Component 2 — No correct weights found in G column")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: H2:H5 contain grade assignment formulas (0.3 points)
    # Each row contributes 0.075 points
    try:
        comp3_score = 0.0
        for row in range(2, 6):
            cell_val = ws.cell(row=row, column=8).value  # H column
            if cell_val is None:
                print(f"FAIL: H{row} is empty — no grade formula")
                continue
            if not isinstance(cell_val, str) or not cell_val.startswith('='):
                print(f"FAIL: H{row} is not a formula: {cell_val}")
                continue

            f = normalize_formula(cell_val)
            # Grade formula should reference G column and have grade thresholds
            has_g_ref = f'G{row}' in f
            has_grades = all(g in f for g in ['"A"', '"B"', '"C"', '"D"'])
            has_thresholds = '80' in f and '60' in f and '40' in f

            if has_g_ref and has_grades and has_thresholds:
                comp3_score += 0.075
                print(f"PASS: H{row} has correct grade formula (refs G{row}, grades A-D, thresholds 80/60/40)")
            elif has_g_ref and has_grades:
                comp3_score += 0.05
                print(f"PARTIAL: H{row} has grade formula with correct grades but missing some thresholds")
            elif has_g_ref:
                comp3_score += 0.025
                print(f"PARTIAL: H{row} references G{row} but missing grade letters")
            else:
                print(f"FAIL: H{row} grade formula missing key elements (G ref: {has_g_ref}, grades: {has_grades}, thresholds: {has_thresholds})")

        if comp3_score > 0:
            print(f"PASS: Component 3 — H column grade formulas ({comp3_score} pts)")
            total_score += comp3_score
        else:
            print(f"FAIL: Component 3 — No valid grade formulas in H column")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
