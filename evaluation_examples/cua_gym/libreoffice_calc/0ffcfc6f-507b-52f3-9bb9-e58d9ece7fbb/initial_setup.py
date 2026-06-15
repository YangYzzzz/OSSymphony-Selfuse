"""
Initial Setup: Lead scoring worksheet with raw data (no formulas)
Task ID: calc_sales_068
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_068'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'LeadScoring'

    # Headers
    headers = ['Lead', 'Company Size', 'Industry Match', 'Budget Confirmed',
               'Decision Timeline', 'Engagement Score', 'Total Score', 'Grade']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_font_white = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows - columns A through F only; G and H left empty for the task
    data = [
        ['Lead 1', 'Enterprise', 'Yes', 'Yes', '<3 months', 85],
        ['Lead 2', 'SMB', 'No', 'No', '>6 months', 30],
        ['Lead 3', 'Mid-Market', 'Yes', 'Yes', '3-6 months', 65],
        ['Lead 4', 'Enterprise', 'Yes', 'No', '<3 months', 72],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name='Calibri', size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set column widths for readability
    col_widths = {'A': 12, 'B': 16, 'C': 16, 'D': 18, 'E': 18, 'F': 18, 'G': 14, 'H': 10}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Add a scoring reference sheet for context
    ws2 = wb.create_sheet('ScoringCriteria')
    ws2['A1'] = 'Criterion'
    ws2['B1'] = 'Value'
    ws2['C1'] = 'Points'
    ws2['A1'].font = Font(bold=True)
    ws2['B1'].font = Font(bold=True)
    ws2['C1'].font = Font(bold=True)

    criteria_data = [
        ['Company Size', 'Enterprise', 30],
        ['Company Size', 'Mid-Market', 20],
        ['Company Size', 'SMB', 10],
        ['Industry Match', 'Yes', 20],
        ['Industry Match', 'No', 0],
        ['Budget Confirmed', 'Yes', 15],
        ['Budget Confirmed', 'No', 0],
        ['Decision Timeline', '<3 months', 20],
        ['Decision Timeline', '3-6 months', 10],
        ['Decision Timeline', '>6 months', 0],
        ['Engagement Score', 'Score/100 * 15', 'Max 15'],
    ]
    for r, row_data in enumerate(criteria_data, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
