"""
Initial Setup: Production capacity planning sheet with raw data (no formulas).
Task ID: calc_ops_057
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_057'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Capacity ---
    ws = wb.active
    ws.title = 'Capacity'

    # Headers
    headers = [
        'Work Center', 'Shifts/Day', 'Hours/Shift', 'Working Days',
        'Available Hours', 'Demand (units)', 'Cycle Time (min)',
        'Required Hours', 'Utilization %'
    ]
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_side = Side(style='thin', color='000000')
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border

    # Data rows (E, H, I columns left EMPTY - those are the task)
    data = [
        # Work Center, Shifts/Day, Hours/Shift, Working Days, (Available Hours=empty),
        # Demand (units), Cycle Time (min), (Required Hours=empty), (Utilization %=empty)
        ['CNC Mill',      2, 8, 22, None, 5000, 4.5, None, None],
        ['Assembly',      3, 8, 22, None, 8000, 6.0, None, None],
        ['Paint Booth',   1, 8, 22, None, 3000, 10.0, None, None],
        ['Weld Station',  2, 8, 22, None, 4000, 8.0, None, None],
    ]

    data_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_align = Alignment(horizontal='center', vertical='center')

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = data_border
            if c >= 2:
                cell.alignment = center_align

    # Column widths for readability
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 17
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Row height for header
    ws.row_dimensions[1].height = 30

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
