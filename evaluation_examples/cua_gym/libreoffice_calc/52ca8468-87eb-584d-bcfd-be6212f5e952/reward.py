"""
Reward Script: Production capacity planning sheet with formulas
Task ID: calc_ops_057
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Available Hours formulas in E2:E5 (=B*C*D)
  Component 2 (0.30): Required Hours formulas in H2:H5 (=F*G/60)
  Component 3 (0.25): Utilization % formulas in I2:I5 (=H/E)
  Component 4 (0.15): Utilization % number format (percentage) in I2:I5
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_057'


def normalize_formula(f):
    """Normalize a formula for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists
    if 'Capacity' not in wb.sheetnames:
        print("FAIL: Sheet 'Capacity' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Capacity']

    # Component 1: Available Hours formulas E2:E5 = B*C*D (0.30 points)
    try:
        e_pass = 0
        for row in range(2, 6):
            val = ws.cell(row=row, column=5).value  # column E
            expected = f'=B{row}*C{row}*D{row}'
            if isinstance(val, str) and normalize_formula(val) == normalize_formula(expected):
                e_pass += 1
            else:
                print(f"FAIL: E{row} expected formula '{expected}', found: {val!r}")
        if e_pass == 4:
            print(f"PASS: Component 1 - All 4 Available Hours formulas correct (0.30 pts)")
            total_score += 0.30
        elif e_pass > 0:
            partial = round(0.30 * e_pass / 4, 2)
            print(f"PARTIAL: Component 1 - {e_pass}/4 Available Hours formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - No Available Hours formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Required Hours formulas H2:H5 = F*G/60 (0.30 points)
    try:
        h_pass = 0
        for row in range(2, 6):
            val = ws.cell(row=row, column=8).value  # column H
            expected = f'=F{row}*G{row}/60'
            if isinstance(val, str) and normalize_formula(val) == normalize_formula(expected):
                h_pass += 1
            else:
                print(f"FAIL: H{row} expected formula '{expected}', found: {val!r}")
        if h_pass == 4:
            print(f"PASS: Component 2 - All 4 Required Hours formulas correct (0.30 pts)")
            total_score += 0.30
        elif h_pass > 0:
            partial = round(0.30 * h_pass / 4, 2)
            print(f"PARTIAL: Component 2 - {h_pass}/4 Required Hours formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 - No Required Hours formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Utilization % formulas I2:I5 = H/E (0.25 points)
    try:
        i_pass = 0
        for row in range(2, 6):
            val = ws.cell(row=row, column=9).value  # column I
            expected = f'=H{row}/E{row}'
            if isinstance(val, str) and normalize_formula(val) == normalize_formula(expected):
                i_pass += 1
            else:
                print(f"FAIL: I{row} expected formula '{expected}', found: {val!r}")
        if i_pass == 4:
            print(f"PASS: Component 3 - All 4 Utilization formulas correct (0.25 pts)")
            total_score += 0.25
        elif i_pass > 0:
            partial = round(0.25 * i_pass / 4, 2)
            print(f"PARTIAL: Component 3 - {i_pass}/4 Utilization formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 - No Utilization formulas found")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Utilization % number format in I2:I5 (0.15 points)
    # The golden file uses '0.0%' format for percentage display
    try:
        fmt_pass = 0
        for row in range(2, 6):
            nf = ws.cell(row=row, column=9).number_format
            # Accept any percentage-style format
            if nf and '%' in str(nf):
                fmt_pass += 1
            else:
                print(f"FAIL: I{row} expected percentage format, found: {nf!r}")
        if fmt_pass == 4:
            print(f"PASS: Component 4 - All 4 Utilization cells have percentage format (0.15 pts)")
            total_score += 0.15
        elif fmt_pass > 0:
            partial = round(0.15 * fmt_pass / 4, 2)
            print(f"PARTIAL: Component 4 - {fmt_pass}/4 cells have percentage format ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 - No percentage formatting found")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
