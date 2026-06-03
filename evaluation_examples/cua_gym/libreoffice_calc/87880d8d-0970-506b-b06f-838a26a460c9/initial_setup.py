"""
Initial Setup: Competitive pricing analysis spreadsheet
Task ID: calc_sales_pricing_competitor_038
Domain: libreoffice_calc

Creates CompetitorPricing sheet with 30 products.
Columns B-E populated with prices; F, G, H are empty (task fills these).
"""

import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_pricing_competitor_038'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'CompetitorPricing'

    # --- Row 1: Headers ---
    headers = ['Product', 'Our Price', 'Competitor A', 'Competitor D', 'Competitor E',
               'Avg Competitor', 'Price Index', 'Risk Flag']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # --- Rows 2-31: 30 products with realistic pricing data ---
    # Our prices range from $299 to $4,999; competitor prices vary ±20%
    products_data = [
        # Product,              Our Price, Comp A,  Comp D,  Comp E
        ('ProScan 3000',           1299,    1199,    1349,    1275),
        ('DataLink Elite',         2499,    2350,    2599,    2450),
        ('CloudSync Pro',           499,     459,     519,     485),
        ('SecureVault X',          3799,    3500,    3899,    3750),
        ('NanoTrack 500',           899,     849,     929,     875),
        ('FlexBoard Ultra',        1599,    1499,    1649,    1580),
        ('IntelliSense Hub',       2199,    2100,    2299,    2150),
        ('QuickDeploy Manager',     699,     649,     729,     685),
        ('StreamFlow 2.0',         1099,    1049,    1149,    1075),
        ('TurboCache Pro',         4999,    4750,    5199,    4850),
        ('LogiTrack Enterprise',   1799,    1699,    1849,    1750),
        ('DataGuard Shield',       2899,    2700,    2999,    2850),
        ('PowerSync 7',             349,     329,     369,     340),
        ('MetroLink 4G',           1499,    1399,    1549,    1475),
        ('VaultSecure 360',        3299,    3100,    3399,    3250),
        ('SmartLink Gateway',       599,     549,     619,     580),
        ('AeroSync Cloud',         2099,    1950,    2199,    2050),
        ('PrecisionTrack 200',      799,     749,     829,     785),
        ('RapidDeploy Plus',       1199,    1099,    1249,    1175),
        ('UltraGuard Pro',         4199,    3900,    4349,    4100),
        ('CoreAnalytics Suite',    1899,    1749,    1949,    1850),
        ('NexGen Router',           449,     419,     469,     440),
        ('HyperLink 8000',         2699,    2500,    2799,    2650),
        ('SkyBridge Connect',       299,     279,     309,     289),
        ('TechVault Premium',      3499,    3250,    3599,    3400),
        ('OmniSync Manager',        999,     929,    1049,     975),
        ('DataPulse Pro',          1699,    1599,    1749,    1675),
        ('MegaTrack 100',          2299,    2150,    2399,    2250),
        ('SecureLink Enterprise',  3999,    3750,    4149,    3900),
        ('FlexCloud Starter',       649,     599,     679,     635),
    ]

    for row_idx, (product, our_price, comp_a, comp_d, comp_e) in enumerate(products_data, 2):
        ws.cell(row=row_idx, column=1, value=product)
        ws.cell(row=row_idx, column=2, value=our_price)
        ws.cell(row=row_idx, column=3, value=comp_a)
        ws.cell(row=row_idx, column=4, value=comp_d)
        ws.cell(row=row_idx, column=5, value=comp_e)
        # Columns F (6), G (7), H (8) intentionally left empty

    # Set column widths for readability
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: CompetitorPricing')
    print(f'  Rows: 1 header + 30 data rows')
    print(f'  Columns A-E populated; F, G, H empty (task fills these)')


create_initial()
