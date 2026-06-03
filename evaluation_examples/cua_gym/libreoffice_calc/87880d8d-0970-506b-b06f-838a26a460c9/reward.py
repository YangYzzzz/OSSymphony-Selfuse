"""
Reward Script: Competitive Pricing Analysis
Task ID: calc_sales_pricing_competitor_038
Domain: libreoffice_calc

Scoring rubric:
  Component 1: AVERAGE formulas in F2:F31          — 0.25 pts
  Component 2: Price index formulas in G2:G31       — 0.25 pts
  Component 3: Risk flag IF formulas in H2:H31      — 0.20 pts
  Component 4: Conditional formatting on G2:G31     — 0.20 pts
  Component 5: Currency format on B2:F31            — 0.10 pts
  Total: 1.00
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_pricing_competitor_038'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # --- Load workbook ---
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # --- Verify sheet exists ---
    if 'CompetitorPricing' not in wb.sheetnames:
        print("CRITICAL: Sheet 'CompetitorPricing' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['CompetitorPricing']

    # -----------------------------------------------------------------------
    # Component 1: AVERAGE formulas in F2:F31 (0.25 points)
    # Each row's F column should contain =AVERAGE(Cx:Ex) referencing cols C-E
    # This was empty in initial file — verifies the agent computed avg competitor price.
    # -----------------------------------------------------------------------
    try:
        avg_formula_count = 0
        avg_formula_correct = 0
        for row in range(2, 32):
            val = ws.cell(row=row, column=6).value  # column F
            if val is not None:
                avg_formula_count += 1
                # Accept =AVERAGE(C{row}:E{row}) case-insensitive
                val_upper = str(val).upper().replace(' ', '')
                expected = f'=AVERAGE(C{row}:E{row})'.upper()
                if val_upper == expected:
                    avg_formula_correct += 1
        if avg_formula_count == 30 and avg_formula_correct == 30:
            print(f"PASS: Component 1 — All 30 AVERAGE formulas in F2:F31 correct (0.25 pts)")
            total_score += 0.25
        elif avg_formula_count == 30 and avg_formula_correct >= 25:
            print(f"PARTIAL: Component 1 — F formulas filled but {30 - avg_formula_correct} incorrect ({avg_formula_correct}/30 correct)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 — Expected 30 AVERAGE formulas in F2:F31, found {avg_formula_count} filled ({avg_formula_correct} correct)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Price index formulas in G2:G31 (0.25 points)
    # Each row's G column should contain =B{row}/F{row}
    # This was empty in initial file — verifies the price index calculation.
    # -----------------------------------------------------------------------
    try:
        pi_formula_count = 0
        pi_formula_correct = 0
        for row in range(2, 32):
            val = ws.cell(row=row, column=7).value  # column G
            if val is not None:
                pi_formula_count += 1
                val_upper = str(val).upper().replace(' ', '')
                expected = f'=B{row}/F{row}'.upper()
                if val_upper == expected:
                    pi_formula_correct += 1
        if pi_formula_count == 30 and pi_formula_correct == 30:
            print(f"PASS: Component 2 — All 30 price index formulas in G2:G31 correct (0.25 pts)")
            total_score += 0.25
        elif pi_formula_count == 30 and pi_formula_correct >= 25:
            print(f"PARTIAL: Component 2 — G formulas filled but {30 - pi_formula_correct} incorrect ({pi_formula_correct}/30 correct)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 — Expected 30 price index formulas in G2:G31, found {pi_formula_count} filled ({pi_formula_correct} correct)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Risk flag IF formulas in H2:H31 (0.20 points)
    # Each row's H column should contain =IF(G{row}>1.15,"At Risk","Competitive")
    # This was empty in initial file — verifies the risk classification.
    # -----------------------------------------------------------------------
    try:
        risk_formula_count = 0
        risk_formula_correct = 0
        for row in range(2, 32):
            val = ws.cell(row=row, column=8).value  # column H
            if val is not None:
                risk_formula_count += 1
                val_upper = str(val).upper().replace(' ', '')
                expected = f'=IF(G{row}>1.15,"AT RISK","COMPETITIVE")'.upper().replace(' ', '')
                if val_upper == expected:
                    risk_formula_correct += 1
        if risk_formula_count == 30 and risk_formula_correct == 30:
            print(f"PASS: Component 3 — All 30 risk flag IF formulas in H2:H31 correct (0.20 pts)")
            total_score += 0.20
        elif risk_formula_count == 30 and risk_formula_correct >= 25:
            print(f"PARTIAL: Component 3 — H formulas filled but {30 - risk_formula_correct} incorrect ({risk_formula_correct}/30 correct)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 3 — Expected 30 IF formulas in H2:H31, found {risk_formula_count} filled ({risk_formula_correct} correct)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: Conditional formatting on G2:G31 (0.20 points)
    # Should have red fill for >1.15 and green fill for <=1.0
    # Initial file had NO conditional formatting — verifies the visual risk highlighting.
    # -----------------------------------------------------------------------
    try:
        has_red_rule = False
        has_green_rule = False
        cf_on_g_col = False

        for cf_range_obj, cf_rules in ws.conditional_formatting._cf_rules.items():
            range_str = str(cf_range_obj).upper().replace(' ', '')
            # Check if the range covers G2:G31 (exact or similar)
            if 'G2' in range_str and 'G31' in range_str:
                cf_on_g_col = True
                for rule in cf_rules:
                    if rule.type == 'cellIs':
                        # Check red fill for >1.15
                        if (rule.operator in ('greaterThan', 'greaterThanOrEqual') and
                                rule.formula and '1.15' in str(rule.formula)):
                            if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                                fill_color = rule.dxf.fill.fgColor.rgb if rule.dxf.fill.fgColor else ''
                                if fill_color.upper().endswith('FF0000'):  # red (e.g. FFFF0000)
                                    has_red_rule = True
                        # Check green fill for <=1.0
                        if (rule.operator in ('lessThanOrEqual', 'lessThan') and
                                rule.formula and '1.0' in str(rule.formula)):
                            if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                                fill_color = rule.dxf.fill.fgColor.rgb if rule.dxf.fill.fgColor else ''
                                if fill_color.upper().endswith('00FF00'):  # green (e.g. FF00FF00)
                                    has_green_rule = True

        if has_red_rule and has_green_rule:
            print(f"PASS: Component 4 — Conditional formatting on G2:G31 correct: red for >1.15, green for <=1.0 (0.20 pts)")
            total_score += 0.20
        elif cf_on_g_col and (has_red_rule or has_green_rule):
            rule_desc = []
            if has_red_rule:
                rule_desc.append("red >1.15")
            if has_green_rule:
                rule_desc.append("green <=1.0")
            print(f"PARTIAL: Component 4 — CF on G2:G31 partially correct: {', '.join(rule_desc)} (0.10 pts)")
            total_score += 0.10
        elif cf_on_g_col:
            print(f"FAIL: Component 4 — CF rule found on G2:G31 but neither red>1.15 nor green<=1.0 detected")
        else:
            print(f"FAIL: Component 4 — No conditional formatting found on G2:G31")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -----------------------------------------------------------------------
    # Component 5: Currency format on B2:F31 and decimal format on G2:G31 (0.10 points)
    # Initial file had 'General' format — verifies number formatting applied.
    # -----------------------------------------------------------------------
    try:
        currency_cols_ok = True
        currency_missing = []
        # Check columns B (2) through F (6) have currency format ($#,##0.00)
        for col in range(2, 7):
            sample_fmt = ws.cell(row=2, column=col).number_format
            if '$' not in str(sample_fmt) and '0.00' not in str(sample_fmt):
                currency_cols_ok = False
                from openpyxl.utils import get_column_letter
                currency_missing.append(f"{get_column_letter(col)}2: {repr(sample_fmt)}")

        g_fmt = ws.cell(row=2, column=7).number_format
        g_format_ok = '0.00' in str(g_fmt)

        if currency_cols_ok and g_format_ok:
            print(f"PASS: Component 5 — Currency format on B-F columns and decimal on G correct (0.10 pts)")
            total_score += 0.10
        elif currency_cols_ok:
            print(f"PARTIAL: Component 5 — Currency format on B-F ok but G format: {repr(g_fmt)} (0.05 pts)")
            total_score += 0.05
        elif g_format_ok:
            print(f"PARTIAL: Component 5 — G decimal format ok but currency missing on: {currency_missing} (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — Missing formats. Currency cols: {currency_missing}, G: {repr(g_fmt)}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
