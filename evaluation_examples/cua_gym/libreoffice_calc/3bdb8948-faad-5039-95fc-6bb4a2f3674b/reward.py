"""
Reward Script: Fill asset owner name and department via VLOOKUP with conditional formatting
Task ID: osworld_calc_vlookup_fill_names_008
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): Column C (Owner Name) contains IFERROR-wrapped VLOOKUP formulas in rows 2-16
  Component 2 (0.35): Column D (Department) contains IFERROR-wrapped VLOOKUP formulas in rows 2-16
  Component 3 (0.30): Conditional formatting on C2:D16 highlights "Not Found" cells in red (cellIs equal)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_vlookup_fill_names_008'

DATA_ROWS = range(2, 17)  # rows 2 through 16 (15 asset rows)


def is_iferror_vlookup(formula):
    """
    Check if the cell formula is an IFERROR-wrapped VLOOKUP.
    Accepts case-insensitive, with 'Not Found' as fallback text.
    """
    if not isinstance(formula, str):
        return False
    clean = formula.upper().replace(' ', '')
    return clean.startswith('=IFERROR(VLOOKUP(') and 'NOTFOUND' in clean.replace('"', '').replace("'", '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Confirm sheet exists
    if 'IT Assets' not in wb.sheetnames:
        print("CRITICAL: Sheet 'IT Assets' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['IT Assets']

    # -------------------------------------------------------------------------
    # Component 1: Column C (Owner Name) has IFERROR-wrapped VLOOKUP formulas
    #              in all data rows (rows 2-16). (0.35 points)
    # -------------------------------------------------------------------------
    try:
        c_formula_count = 0
        c_missing = []
        for row in DATA_ROWS:
            val = ws.cell(row=row, column=3).value  # column C
            if is_iferror_vlookup(val):
                c_formula_count += 1
            else:
                c_missing.append(f"C{row}={repr(val)}")

        expected = len(DATA_ROWS)
        if c_formula_count == expected:
            print(f"PASS: Component 1 — All {expected} cells in column C contain IFERROR(VLOOKUP) formulas (0.35 pts)")
            total_score += 0.35
        elif c_formula_count > 0:
            partial = round(0.35 * c_formula_count / expected, 4)
            print(f"PARTIAL: Component 1 — {c_formula_count}/{expected} column C cells have IFERROR(VLOOKUP) formulas; "
                  f"missing: {c_missing[:3]}... ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No IFERROR(VLOOKUP) formulas found in column C; "
                  f"sample: {c_missing[:3]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Column D (Department) has IFERROR-wrapped VLOOKUP formulas
    #              in all data rows (rows 2-16). (0.35 points)
    # -------------------------------------------------------------------------
    try:
        d_formula_count = 0
        d_missing = []
        for row in DATA_ROWS:
            val = ws.cell(row=row, column=4).value  # column D
            if is_iferror_vlookup(val):
                d_formula_count += 1
            else:
                d_missing.append(f"D{row}={repr(val)}")

        expected = len(DATA_ROWS)
        if d_formula_count == expected:
            print(f"PASS: Component 2 — All {expected} cells in column D contain IFERROR(VLOOKUP) formulas (0.35 pts)")
            total_score += 0.35
        elif d_formula_count > 0:
            partial = round(0.35 * d_formula_count / expected, 4)
            print(f"PARTIAL: Component 2 — {d_formula_count}/{expected} column D cells have IFERROR(VLOOKUP) formulas; "
                  f"missing: {d_missing[:3]}... ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No IFERROR(VLOOKUP) formulas found in column D; "
                  f"sample: {d_missing[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Conditional formatting on C/D range highlights "Not Found"
    #              cells in red (cellIs equal, fill FFFF0000). (0.30 points)
    # -------------------------------------------------------------------------
    try:
        cf_pass = False
        cf_details = []
        for cf_range in ws.conditional_formatting:
            # cf_range is a string like "C2:D16" or similar
            cf_range_str = str(cf_range)
            for rule in ws.conditional_formatting[cf_range]:
                rule_type = getattr(rule, 'type', '')
                operator = getattr(rule, 'operator', '')
                formula = getattr(rule, 'formula', [])

                # Must be a cellIs rule with operator "equal" targeting "Not Found"
                is_cell_is_equal = (rule_type == 'cellIs' and operator == 'equal')
                # formula is a list like ['"Not Found"']; check case-insensitively
                has_not_found = any(
                    'not found' in str(f).lower() for f in formula
                ) if formula else False

                # Check fill color is red-ish (FFFF0000)
                has_red_fill = False
                dxf = getattr(rule, 'dxf', None)
                if dxf and hasattr(dxf, 'fill') and dxf.fill:
                    try:
                        rgb = dxf.fill.fgColor.rgb
                        # Accept pure red: FFFF0000 or FF0000 (6-char)
                        has_red_fill = rgb in ('FFFF0000', '00FF0000', 'FF0000')
                    except Exception:
                        pass

                # The CF range should cover columns C and D (columns 3-4)
                covers_c_d = ('C' in cf_range_str and 'D' in cf_range_str) or \
                             re.search(r'C\d', cf_range_str) is not None

                cf_details.append({
                    'range': cf_range_str,
                    'type': rule_type,
                    'operator': operator,
                    'formula': formula,
                    'has_red_fill': has_red_fill,
                    'has_not_found': has_not_found,
                    'covers_c_d': covers_c_d
                })

                if is_cell_is_equal and has_not_found and has_red_fill:
                    cf_pass = True

        if cf_pass:
            print(f"PASS: Component 3 — Conditional formatting found on C/D range: "
                  f"cellIs equal 'Not Found' with red fill (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 3 — Required conditional formatting not found. "
                  f"Details: {cf_details}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
