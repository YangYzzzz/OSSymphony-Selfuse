"""
Initial Setup: IT Asset Tracker with Employee Directory (VLOOKUP task)
Task ID: osworld_calc_vlookup_fill_names_008
Domain: libreoffice_calc

Creates an IT asset spreadsheet where:
- Columns A-E contain asset records (C and D are empty — to be filled with VLOOKUP)
- Columns G-I contain an employee directory (User ID → Name → Department)
- Some user IDs in column B do NOT exist in the directory (to trigger IFERROR)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_vlookup_fill_names_008'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: IT Assets ---
    ws = wb.active
    ws.title = "IT Assets"

    # ---- Asset table headers (A1:E1) ----
    asset_headers = ["Asset ID", "Assigned User ID", "Owner Name", "Department", "Asset Type"]
    for col, h in enumerate(asset_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # ---- Asset data rows (rows 2–16) ----
    # Note: Columns C (Owner Name) and D (Department) are LEFT EMPTY intentionally.
    # Some user IDs (e.g., U007, U012, U015) do NOT exist in the directory.
    asset_data = [
        # AssetID, UserID,  OwnerName(empty), Dept(empty), AssetType
        ["AST-001", "U001", None, None, "Laptop"],
        ["AST-002", "U002", None, None, "Monitor"],
        ["AST-003", "U003", None, None, "Keyboard"],
        ["AST-004", "U004", None, None, "Laptop"],
        ["AST-005", "U005", None, None, "Printer"],
        ["AST-006", "U006", None, None, "Laptop"],
        ["AST-007", "U007", None, None, "Docking Station"],   # U007 NOT in directory
        ["AST-008", "U008", None, None, "Headset"],
        ["AST-009", "U009", None, None, "Webcam"],
        ["AST-010", "U010", None, None, "Laptop"],
        ["AST-011", "U011", None, None, "Monitor"],
        ["AST-012", "U012", None, None, "Tablet"],            # U012 NOT in directory
        ["AST-013", "U013", None, None, "Laptop"],
        ["AST-014", "U014", None, None, "Mouse"],
        ["AST-015", "U015", None, None, "Laptop"],            # U015 NOT in directory
    ]

    for r, row_data in enumerate(asset_data, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        ws.cell(row=r, column=2, value=row_data[1])
        # Columns C and D are intentionally empty (None)
        ws.cell(row=r, column=3, value=row_data[2])
        ws.cell(row=r, column=4, value=row_data[3])
        ws.cell(row=r, column=5, value=row_data[4])

    # ---- Separator (column F is empty) ----

    # ---- Employee Directory headers (G1:I1) ----
    dir_headers = ["User ID", "Full Name", "Department"]
    for col, h in enumerate(dir_headers, 7):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill(start_color="FF217346", end_color="FF217346", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # ---- Employee Directory data (rows 2–13, 12 employees) ----
    # U007, U012, U015 are intentionally omitted so VLOOKUP returns an error.
    directory_data = [
        ["U001", "Sarah Chen",        "Engineering"],
        ["U002", "Marcus Johnson",    "Marketing"],
        ["U003", "Priya Patel",       "Finance"],
        ["U004", "James O'Brien",     "Engineering"],
        ["U005", "Aisha Williams",    "Operations"],
        ["U006", "Lucas Ferreira",    "HR"],
        ["U008", "Nina Kowalski",     "Engineering"],
        ["U009", "Daniel Park",       "IT Support"],
        ["U010", "Elena Vasquez",     "Sales"],
        ["U011", "Omar Hassan",       "Finance"],
        ["U013", "Chloe Dupont",      "Marketing"],
        ["U014", "Ravi Shankar",      "Engineering"],
    ]

    for r, row_data in enumerate(directory_data, 2):
        ws.cell(row=r, column=7, value=row_data[0])
        ws.cell(row=r, column=8, value=row_data[1])
        ws.cell(row=r, column=9, value=row_data[2])

    # ---- Column widths for readability ----
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 4
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 22
    ws.column_dimensions["I"].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
