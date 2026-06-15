"""
Initial Setup: Open Chrome and LibreOffice Calc with agent_papers.ods (empty data, summary structure)
Task ID: osworld_multi_apps_arxiv_llms_calc_008
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_arxiv_llms_calc_008'
XLSX_TMP = f'{WORKDIR}/{TASK_ID}_tmp.xlsx'
OUTPUT = f'{WORKDIR}/agent_papers.ods'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # --- Main data headers in row 1 ---
    # Columns A-D: arXiv ID, Title, Authors, Keyword
    headers_main = ['arXiv ID', 'Title', 'Authors', 'Keyword']
    for col, h in enumerate(headers_main, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # --- Summary table headers: F1:G1 = 'Keyword' / 'Count' ---
    ws.cell(row=1, column=6, value='Keyword').font = Font(bold=True)
    ws.cell(row=1, column=7, value='Count').font = Font(bold=True)

    # --- F2:F4 = keyword labels ---
    ws.cell(row=2, column=6, value='reasoning')
    ws.cell(row=3, column=6, value='planning')
    ws.cell(row=4, column=6, value='agent')

    # --- G2:G4 are EMPTY (agent must add COUNTIF formulas) ---

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 20  # arXiv ID
    ws.column_dimensions['B'].width = 60  # Title
    ws.column_dimensions['C'].width = 40  # Authors
    ws.column_dimensions['D'].width = 15  # Keyword
    ws.column_dimensions['E'].width = 5   # spacer
    ws.column_dimensions['F'].width = 15  # Keyword (summary)
    ws.column_dimensions['G'].width = 10  # Count

    # Save as .xlsx first, then convert to .ods using LibreOffice
    wb.save(XLSX_TMP)
    print(f'Temporary xlsx created: {XLSX_TMP}')

    # Convert xlsx -> ods using LibreOffice headless
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    result = subprocess.run(
        ['libreoffice', '--headless', '--convert-to', 'ods',
         '--outdir', WORKDIR, XLSX_TMP],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        env=env,
        timeout=60,
    )
    # The output file will be named after the input base name with .ods extension
    converted = f'{WORKDIR}/{TASK_ID}_tmp.ods'
    if os.path.exists(converted):
        os.rename(converted, OUTPUT)
        print(f'Converted and renamed to: {OUTPUT}')
    else:
        print(f'Conversion output: {result.stdout.decode()} {result.stderr.decode()}')
        # Fallback: save directly as .ods using xlsx format (LibreOffice accepts it)
        wb.save(OUTPUT)
        print(f'Fallback: saved directly as {OUTPUT}')

    # Remove temp xlsx
    if os.path.exists(XLSX_TMP):
        os.remove(XLSX_TMP)

    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open Chrome and LibreOffice Calc
    # Kill any existing LibreOffice instances to avoid conflicts
    subprocess.run(['pkill', '-f', 'soffice'], stdout=subprocess.DEVNULL,
                   stderr=subprocess.DEVNULL)
    time.sleep(1.5)

    # Open Chrome (non-blocking)
    launch_gui('google-chrome --new-window', delay_sec=2.0)

    # Open LibreOffice Calc with the .ods file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)

    print('GUI_READY: launched Chrome and LibreOffice Calc with DISPLAY=:0')


create_initial()
