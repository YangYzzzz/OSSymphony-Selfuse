"""
Initial Setup: Create a project plan spreadsheet for Gantt chart task
Task ID: calc_gcp_047
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_047'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: ProjectPlan ---
    ws = wb.active
    ws.title = 'ProjectPlan'

    # Headers
    headers = ['Task', 'StartDay', 'Duration', 'EndDay']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font_white = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    tasks = [
        ('Requirements',  1,  7),
        ('UI Design',     8,  10),
        ('Backend Dev',   12, 18),
        ('Frontend Dev',  15, 15),
        ('Integration',   25, 8),
        ('Testing',       30, 10),
        ('UAT',           38, 5),
        ('Deployment',    42, 3),
    ]

    data_font = Font(name='Calibri', size=11)
    for r, (task_name, start_day, duration) in enumerate(tasks, 2):
        end_day = start_day + duration - 1

        cell_a = ws.cell(row=r, column=1, value=task_name)
        cell_a.font = data_font
        cell_a.border = thin_border

        cell_b = ws.cell(row=r, column=2, value=start_day)
        cell_b.font = data_font
        cell_b.alignment = Alignment(horizontal='center')
        cell_b.border = thin_border

        cell_c = ws.cell(row=r, column=3, value=duration)
        cell_c.font = data_font
        cell_c.alignment = Alignment(horizontal='center')
        cell_c.border = thin_border

        cell_d = ws.cell(row=r, column=4, value=end_day)
        cell_d.font = data_font
        cell_d.alignment = Alignment(horizontal='center')
        cell_d.border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12

    # NO chart in initial state - the task is to create the Gantt chart

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
