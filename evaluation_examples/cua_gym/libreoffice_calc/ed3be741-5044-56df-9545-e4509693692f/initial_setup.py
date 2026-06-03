"""
Initial Setup: Unmerge title cells and sort data
Task ID: calc_tbl_021
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_021'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Data"

    # --- Title: Merge A1:F3 ---
    ws.merge_cells("A1:F3")
    ws["A1"] = "Annual Sales Report"
    ws["A1"].font = Font(name="Arial", size=18, bold=True, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")

    # --- Row 4: spacer (empty) ---

    # --- Row 5: Headers ---
    headers = ["Name", "Region", "Q1 Sales", "Q2 Sales", "Q3 Sales", "Q4 Sales"]
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Column widths ---
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14

    # --- Data rows 6-100 (95 rows) ---
    first_names = [
        "Sarah", "Marcus", "Elena", "James", "Priya", "David", "Mei",
        "Carlos", "Fatima", "Oliver", "Yuki", "Benjamin", "Aisha",
        "Liam", "Sofia", "Ethan", "Nadia", "Ryan", "Amara", "Tyler",
        "Ingrid", "Hassan", "Chloe", "Raj", "Natalie", "Kenji",
        "Isabelle", "Andre", "Zara", "Patrick", "Lucia", "Derek",
    ]
    last_names = [
        "Chen", "Johnson", "Petrov", "Williams", "Sharma", "Kim",
        "Rodriguez", "Ali", "Martinez", "Lee", "Tanaka", "Foster",
        "Okafor", "Nguyen", "Anderson", "Patel", "Brown", "Garcia",
        "Singh", "Wilson", "Johansson", "Mueller", "Dubois", "Park",
        "Clark", "Torres", "Scott", "Wright", "Adams", "Lopez",
    ]
    # Deliberately NOT alphabetical regions to ensure unsorted by column B
    regions = [
        "West", "Northeast", "Midwest", "Southeast", "Southwest",
        "Pacific", "Central", "Northeast", "West", "Southeast",
        "Midwest", "Pacific", "Central", "Southwest", "Northeast",
    ]

    random.seed(42)  # reproducible

    data_rows = []
    for i in range(95):
        fname = first_names[i % len(first_names)]
        lname = last_names[i % len(last_names)]
        name = f"{fname} {lname}"
        region = regions[i % len(regions)]
        q1 = round(random.uniform(15000, 95000), 2)
        q2 = round(random.uniform(15000, 95000), 2)
        q3 = round(random.uniform(15000, 95000), 2)
        q4 = round(random.uniform(15000, 95000), 2)
        data_rows.append([name, region, q1, q2, q3, q4])

    # Shuffle to ensure data is NOT sorted by region (column B)
    random.shuffle(data_rows)

    currency_fmt = '$#,##0.00'
    for r, row_data in enumerate(data_rows, 6):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c >= 3:  # Q1-Q4 columns
                cell.number_format = currency_fmt
            if c <= 2:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
