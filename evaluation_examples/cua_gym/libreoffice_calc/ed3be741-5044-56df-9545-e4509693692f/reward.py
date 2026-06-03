"""
Reward Script: Unmerge title cells and sort data by column B ascending
Task ID: calc_tbl_021
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Title cells A1:F3 are unmerged
  Component 2 (0.6): Data in A6:F100 sorted by column B (Region) ascending
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_021'


def persist_app_state(domain: str):
    """Attempt to save any unsaved GUI edits via Ctrl+S."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Component 1: Title cells A1:F3 are unmerged (0.4 points)
    # In the initial file, A1:F3 is a single merged range.
    # After the task, there should be NO merged cells covering A1:F3.
    try:
        merged_ranges = list(ws.merged_cells.ranges)
        # Check if any merged range overlaps with A1:F3
        overlapping = [
            mr for mr in merged_ranges
            if mr.min_row <= 3 and mr.max_row >= 1 and mr.min_col <= 6 and mr.max_col >= 1
        ]

        if len(overlapping) == 0:
            print(f"PASS: Component 1 -- Title cells A1:F3 are unmerged. Merged ranges: {merged_ranges} (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 -- Title cells A1:F3 still merged. Overlapping ranges: {overlapping}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Data in rows 6-100 sorted by column B (Region) ascending (0.6 points)
    # In the initial file, column B data is not sorted.
    # After the task, column B values should be in ascending alphabetical order.
    try:
        col_b_values = []
        for r in range(6, 101):
            val = ws.cell(row=r, column=2).value
            if val is not None:
                col_b_values.append(str(val).strip())

        if len(col_b_values) < 2:
            print(f"FAIL: Component 2 -- Not enough data rows found (got {len(col_b_values)})")
        else:
            # Check if sorted ascending (case-insensitive)
            is_sorted = all(
                col_b_values[i].lower() <= col_b_values[i + 1].lower()
                for i in range(len(col_b_values) - 1)
            )
            if is_sorted:
                print(f"PASS: Component 2 -- Data sorted by column B ascending. {len(col_b_values)} rows verified. (0.6 pts)")
                total_score += 0.6
            else:
                # Find first out-of-order pair for diagnostics
                for i in range(len(col_b_values) - 1):
                    if col_b_values[i].lower() > col_b_values[i + 1].lower():
                        print(f"FAIL: Component 2 -- Data NOT sorted by column B. First violation at data row {i}: '{col_b_values[i]}' > '{col_b_values[i+1]}'")
                        break
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
