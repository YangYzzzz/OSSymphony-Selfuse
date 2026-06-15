"""
Reward Script: HR Leave Calendar Cleanup
Task ID: calc_hr_leave_calendar_023
Domain: libreoffice_calc
Scoring:
  - Component 1: Duplicate rows removed (row count = 156 data rows + 1 header) — 0.30 pts
  - Component 2: Data sorted by Start Date (col E) in ascending order — 0.30 pts
  - Component 3: Column G has =Fn-En+1 formula for every data row — 0.40 pts
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_hr_leave_calendar_023'
SHEET_NAME = 'Leave Requests'

# Expected constants from task context
INITIAL_DATA_ROWS = 186       # rows 2-187 in the initial file
EXPECTED_DUPLICATE_COUNT = 30 # duplicates to remove
EXPECTED_DATA_ROWS = 156      # 186 - 30 = 156 data rows after dedup (total rows = 157 with header)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition gate: must be able to load the file
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: the sheet must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found in workbook. Found: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Precondition gate: headers must be intact in row 1
    expected_headers = ['Request ID', 'Emp ID', 'Employee Name', 'Leave Type', 'Start Date', 'End Date', 'Total Days']
    actual_headers = [ws.cell(row=1, column=c).value for c in range(1, 8)]
    if actual_headers != expected_headers:
        print(f"CRITICAL: Headers don't match expected. Found: {actual_headers}")
        print("REWARD: 0.0")
        return 0.0

    # ----- Component 1: Correct row count after duplicate removal (0.30 points) -----
    # Task asks to remove duplicate rows where (Emp ID, Start Date, End Date) match.
    # Initial had 186 data rows with 30 duplicates → golden should have 156 data rows.
    # This FAILS on initial (186 rows) and PASSES on golden (156 rows).
    try:
        actual_data_rows = ws.max_row - 1  # subtract header row
        if actual_data_rows == EXPECTED_DATA_ROWS:
            print(f"PASS: Component 1 — Row count = {actual_data_rows} data rows (expected {EXPECTED_DATA_ROWS}, {EXPECTED_DUPLICATE_COUNT} duplicates removed) (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 — Expected {EXPECTED_DATA_ROWS} data rows after removing {EXPECTED_DUPLICATE_COUNT} duplicates, found {actual_data_rows}")
            # Partial check: are there fewer rows than the initial (some dedup was done)?
            if actual_data_rows < INITIAL_DATA_ROWS and actual_data_rows > EXPECTED_DATA_ROWS:
                print(f"  (partial dedup detected: {INITIAL_DATA_ROWS - actual_data_rows} duplicates removed, expected {EXPECTED_DUPLICATE_COUNT})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ----- Component 2: Data sorted by Start Date (col E) ascending (0.30 points) -----
    # Task asks to sort remaining records by start date.
    # This FAILS on initial (unsorted) and PASSES on golden (sorted).
    try:
        dates = []
        for row in range(2, ws.max_row + 1):
            e_val = ws.cell(row=row, column=5).value
            dates.append(e_val)

        # Find first violation of ascending order
        first_violation = None
        for i in range(len(dates) - 1):
            d_curr = dates[i]
            d_next = dates[i + 1]
            if d_curr is not None and d_next is not None:
                if d_curr > d_next:
                    first_violation = (i + 2, i + 3, d_curr, d_next)
                    break

        if first_violation is None:
            print(f"PASS: Component 2 — Data is sorted by Start Date (col E) in ascending order, first={dates[0]}, last={dates[-1]} (0.30 pts)")
            total_score += 0.30
        else:
            row_a, row_b, d_a, d_b = first_violation
            print(f"FAIL: Component 2 — Data not sorted by Start Date. First violation at rows {row_a}/{row_b}: {d_a} > {d_b}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ----- Component 3: Column G has =Fn-En+1 formula for every data row (0.40 points) -----
    # Task asks to add a formula =F{n}-E{n}+1 for calendar days (inclusive) in col G.
    # This FAILS on initial (G column is empty) and PASSES on golden (all rows have formula).
    try:
        max_data_row = ws.max_row
        formula_count = 0
        total_data_rows = max_data_row - 1  # rows 2 through max_row
        formula_errors = []

        for row in range(2, max_data_row + 1):
            g_val = ws.cell(row=row, column=7).value
            if g_val is None:
                formula_errors.append(f"Row {row}: G is None (empty)")
                continue

            g_str = str(g_val).strip()
            # Accept =Fn-En+1 pattern (case insensitive, any row number)
            # Pattern: =F<row>-E<row>+1
            expected_formula = f"=F{row}-E{row}+1"
            # Normalize for comparison (uppercase, no spaces)
            g_normalized = g_str.upper().replace(" ", "")
            expected_normalized = expected_formula.upper().replace(" ", "")

            if g_normalized == expected_normalized:
                formula_count += 1
            else:
                formula_errors.append(f"Row {row}: expected '{expected_formula}', found '{g_val}'")

        if formula_count == total_data_rows and total_data_rows > 0:
            print(f"PASS: Component 3 — All {formula_count}/{total_data_rows} data rows have =Fn-En+1 formula in col G (0.40 pts)")
            total_score += 0.40
        elif formula_count > 0:
            ratio = formula_count / total_data_rows
            print(f"FAIL: Component 3 — Only {formula_count}/{total_data_rows} rows have correct =Fn-En+1 formula.")
            if formula_errors[:3]:
                print(f"  First errors: {formula_errors[:3]}")
        else:
            print(f"FAIL: Component 3 — No rows have =Fn-En+1 formula in col G. Column G appears to be empty.")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
