"""
Initial Setup: HR Leave Request Log with duplicates, unsorted, empty Total Days column
Task ID: calc_hr_leave_calendar_023
Domain: libreoffice_calc
"""

import os
from datetime import date, timedelta
import random
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_leave_calendar_023'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Leave Requests'

    # Headers: A='Request ID', B='Emp ID', C='Employee Name', D='Leave Type', E='Start Date', F='End Date', G='Total Days'
    headers = ['Request ID', 'Emp ID', 'Employee Name', 'Leave Type', 'Start Date', 'End Date', 'Total Days']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Employee pool
    employees = [
        ('E001', 'Sarah Chen'),
        ('E002', 'Marcus Johnson'),
        ('E003', 'Priya Patel'),
        ('E004', 'David Okonkwo'),
        ('E005', 'Yuki Tanaka'),
        ('E006', 'Fatima Al-Rashid'),
        ('E007', 'James O\'Brien'),
        ('E008', 'Luciana Ferreira'),
        ('E009', 'Aleksandr Novikov'),
        ('E010', 'Mei-Ling Zhou'),
        ('E011', 'Carlos Rivera'),
        ('E012', 'Amina Hassan'),
        ('E013', 'Thomas Mueller'),
        ('E014', 'Nadia Kowalski'),
        ('E015', 'Rajesh Iyer'),
    ]

    leave_types = ['Annual Leave', 'Sick Leave', 'Maternity Leave', 'Paternity Leave',
                   'Bereavement Leave', 'Unpaid Leave', 'Compensatory Leave', 'Study Leave']

    base_date = date(2024, 1, 1)

    # Generate unique leave records first (without duplicates)
    unique_records = []
    req_id = 1001

    for _ in range(150):
        emp_id, emp_name = random.choice(employees)
        leave_type = random.choice(leave_types)
        # Random start date in 2024-2025
        day_offset = random.randint(0, 540)  # ~18 months
        start_dt = base_date + timedelta(days=day_offset)
        duration = random.randint(1, 14)
        end_dt = start_dt + timedelta(days=duration - 1)
        unique_records.append({
            'req_id': f'REQ-{req_id:04d}',
            'emp_id': emp_id,
            'emp_name': emp_name,
            'leave_type': leave_type,
            'start': start_dt,
            'end': end_dt,
        })
        req_id += 1

    # Select ~30 records to duplicate (same emp_id, start, end — different req_id allowed)
    dup_indices = random.sample(range(len(unique_records)), 30)
    duplicates = []
    for idx in dup_indices:
        orig = unique_records[idx]
        duplicates.append({
            'req_id': f'REQ-{req_id:04d}',
            'emp_id': orig['emp_id'],
            'emp_name': orig['emp_name'],
            'leave_type': orig['leave_type'],
            'start': orig['start'],
            'end': orig['end'],
        })
        req_id += 1

    # Combine: 150 unique + 36 extras that include 30 duplicates = 186 total records
    # Add a few more unique records to reach 186
    for _ in range(6):
        emp_id, emp_name = random.choice(employees)
        leave_type = random.choice(leave_types)
        day_offset = random.randint(0, 540)
        start_dt = base_date + timedelta(days=day_offset)
        duration = random.randint(1, 10)
        end_dt = start_dt + timedelta(days=duration - 1)
        unique_records.append({
            'req_id': f'REQ-{req_id:04d}',
            'emp_id': emp_id,
            'emp_name': emp_name,
            'leave_type': leave_type,
            'start': start_dt,
            'end': end_dt,
        })
        req_id += 1

    all_records = unique_records + duplicates  # 156 + 30 = 186 records

    # Shuffle the combined records (so they're not sorted, and duplicates are mixed in)
    random.shuffle(all_records)

    # Write to worksheet (rows 2-187, column G left empty)
    for r, rec in enumerate(all_records, 2):
        ws.cell(row=r, column=1, value=rec['req_id'])
        ws.cell(row=r, column=2, value=rec['emp_id'])
        ws.cell(row=r, column=3, value=rec['emp_name'])
        ws.cell(row=r, column=4, value=rec['leave_type'])
        ws.cell(row=r, column=5, value=rec['start'])
        ws.cell(row=r, column=5).number_format = 'yyyy-mm-dd'
        ws.cell(row=r, column=6, value=rec['end'])
        ws.cell(row=r, column=6).number_format = 'yyyy-mm-dd'
        # Column G (Total Days) intentionally left empty

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Total data rows: {len(all_records)}')
    print(f'Duplicate records added: {len(duplicates)}')

create_initial()
