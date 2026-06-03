"""
reward.py – Loan Amortization Schedule Verifier (calc_wf_014)

Scoring rubric (6 components):
  1. PMT formula in B5                              ~0.15
  2. Amortization table has 360 rows of data        ~0.25
  3. Interest calculation correct                    ~0.15
  4. Principal calculation correct                   ~0.15
  5. Chart exists with 2 data series                 ~0.15
  6. Sheet protection enabled + formula cells locked ~0.15
"""

import openpyxl
import re

FILE_PATH = "/home/user/calc_wf_014.xlsx"
SHEET_NAME = "Loan Calculator"

def load_workbook_safe(path):
    try:
        return openpyxl.load_workbook(path)
    except Exception:
        return None

def check_pmt_formula(ws):
    """Check B5 contains a PMT formula referencing rate, term, principal."""
    val = ws["B5"].value
    if not isinstance(val, str):
        return 0.0
    val_upper = val.upper().replace(" ", "")
    # Must contain PMT function
    if "PMT(" not in val_upper:
        return 0.0
    # Should reference B3 (rate), B4 (term), B2 (principal)
    has_rate = "B3" in val_upper
    has_term = "B4" in val_upper
    has_principal = "B2" in val_upper
    if has_rate and has_term and has_principal:
        return 1.0
    # Partial credit if PMT exists but references are off
    return 0.5

def check_table_rows(ws):
    """Check amortization table has ~360 rows of data starting from row 8."""
    # Headers should be in row 7
    header_a7 = ws.cell(row=7, column=1).value
    if header_a7 is None or "payment" not in str(header_a7).lower():
        return 0.0

    # Count data rows: rows 8+ where column A (Payment #) has a value
    data_row_count = 0
    for row in range(8, 400):
        val = ws.cell(row=row, column=1).value
        if val is not None:
            data_row_count += 1
        else:
            # Allow small gaps but stop after a big gap
            next_val = ws.cell(row=row + 1, column=1).value if row < 399 else None
            if next_val is None:
                break

    if data_row_count >= 360:
        return 1.0
    elif data_row_count >= 300:
        return 0.7
    elif data_row_count >= 100:
        return 0.4
    elif data_row_count >= 10:
        return 0.2
    return 0.0

def check_interest_formulas(ws):
    """Check interest calculation: balance * monthly rate pattern."""
    # Row 8 (first data row): Interest = Principal * (Rate/12)
    c8 = ws.cell(row=8, column=3).value  # C8
    if not isinstance(c8, str):
        return 0.0

    c8_upper = c8.upper().replace(" ", "").replace("$", "")
    # Should reference B2 (principal) and B3 (rate) with /12
    first_row_ok = ("B2" in c8_upper and "B3" in c8_upper) or ("B2" in c8_upper and "/12" in c8_upper)

    # Row 9+: Interest = previous_balance * (rate/12)
    c9 = ws.cell(row=9, column=3).value  # C9
    if not isinstance(c9, str):
        return 0.5 if first_row_ok else 0.0

    c9_upper = c9.upper().replace(" ", "").replace("$", "")
    # Should reference E8 (previous balance) and B3 (rate)
    subsequent_ok = ("E8" in c9_upper and "B3" in c9_upper)

    if first_row_ok and subsequent_ok:
        return 1.0
    elif first_row_ok or subsequent_ok:
        return 0.5
    return 0.0

def check_principal_formulas(ws):
    """Check principal calculation: payment - interest pattern."""
    d8 = ws.cell(row=8, column=4).value  # D8
    if not isinstance(d8, str):
        return 0.0

    d8_upper = d8.upper().replace(" ", "").replace("$", "")
    # Should be =B8-C8 pattern (payment minus interest)
    has_payment_ref = "B8" in d8_upper
    has_interest_ref = "C8" in d8_upper
    has_subtraction = "-" in d8_upper

    first_ok = has_payment_ref and has_interest_ref and has_subtraction

    # Also check balance formula E8
    e8 = ws.cell(row=8, column=5).value
    balance_ok = False
    if isinstance(e8, str):
        e8_upper = e8.upper().replace(" ", "").replace("$", "")
        # Should reference B2 (initial principal) and D8 (principal portion)
        balance_ok = "D8" in e8_upper and ("B2" in e8_upper or "E7" in e8_upper)

    if first_ok and balance_ok:
        return 1.0
    elif first_ok or balance_ok:
        return 0.5
    return 0.0

def check_chart(ws):
    """Check that a chart exists with 2 data series (Interest and Principal)."""
    charts = ws._charts
    if len(charts) == 0:
        return 0.0

    best_score = 0.0
    for chart in charts:
        series_count = len(chart.series)
        # Check if it's a line chart (preferred) or any chart type
        is_line = isinstance(chart, openpyxl.chart.LineChart)

        if series_count >= 2:
            # Check series reference Interest (col C) and Principal (col D)
            series_refs = []
            for s in chart.series:
                if hasattr(s, 'title') and s.title is not None:
                    title_obj = s.title
                    if hasattr(title_obj, 'strRef') and title_obj.strRef is not None:
                        ref_str = str(title_obj.strRef.f) if hasattr(title_obj.strRef, 'f') else ""
                        series_refs.append(ref_str.upper())

            has_interest = any("C7" in r or "C1" in r for r in series_refs)
            has_principal = any("D7" in r or "D1" in r for r in series_refs)

            if has_interest and has_principal:
                score = 1.0
            elif series_count >= 2:
                score = 0.7
            else:
                score = 0.4

            if is_line:
                best_score = max(best_score, score)
            else:
                best_score = max(best_score, score * 0.8)
        elif series_count == 1:
            best_score = max(best_score, 0.3)

    return best_score

def check_protection(ws):
    """Check sheet protection enabled and formula cells are locked while input cells are unlocked."""
    # Sheet protection must be enabled
    if not ws.protection.sheet:
        return 0.0

    score = 0.5  # Protection is enabled

    # Input cells (B2, B3, B4) should be unlocked
    input_unlocked = True
    for coord in ["B2", "B3", "B4"]:
        cell = ws[coord]
        if cell.protection.locked is None or cell.protection.locked:
            input_unlocked = False
            break

    # Formula cells should be locked (default)
    formula_locked = True
    for coord in ["B5", "B8", "C8", "D8", "E8"]:
        cell = ws[coord]
        # locked=True or locked=None (default=locked)
        if cell.protection.locked is not None and not cell.protection.locked:
            formula_locked = False
            break

    if input_unlocked and formula_locked:
        score = 1.0
    elif formula_locked:
        score = 0.7
    elif input_unlocked:
        score = 0.7

    return score


def main():
    wb = load_workbook_safe(FILE_PATH)
    if wb is None:
        print("ERROR: Cannot load workbook")
        print("REWARD: 0.0")
        return

    if SHEET_NAME not in wb.sheetnames:
        # Try first sheet
        ws = wb.worksheets[0]
        if ws.title.lower().replace(" ", "") != SHEET_NAME.lower().replace(" ", ""):
            print(f"WARNING: Sheet '{SHEET_NAME}' not found, using '{ws.title}'")
    else:
        ws = wb[SHEET_NAME]

    weights = {
        "pmt_formula": 0.15,
        "table_rows": 0.25,
        "interest_calc": 0.15,
        "principal_calc": 0.15,
        "chart": 0.15,
        "protection": 0.15,
    }

    scores = {}
    scores["pmt_formula"] = check_pmt_formula(ws)
    scores["table_rows"] = check_table_rows(ws)
    scores["interest_calc"] = check_interest_formulas(ws)
    scores["principal_calc"] = check_principal_formulas(ws)
    scores["chart"] = check_chart(ws)
    scores["protection"] = check_protection(ws)

    total = 0.0
    for key, weight in weights.items():
        component = scores[key] * weight
        total += component
        print(f"  {key}: {scores[key]:.2f} x {weight} = {component:.3f}")

    # Round to 1 decimal
    total = round(total, 1)
    # Clamp
    total = max(0.0, min(1.0, total))

    print(f"REWARD: {total}")


if __name__ == "__main__":
    main()
