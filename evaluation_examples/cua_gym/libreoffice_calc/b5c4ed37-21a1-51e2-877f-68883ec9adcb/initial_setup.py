"""
initial_setup.py - Create the pre-task state for calc_wf_014.
Creates a workbook with:
- Sheet 'Loan Calculator'
- Input section (rows 1-4) with labels and values
- Amortization table headers at row 7
- NO formulas, NO chart, NO cell protection
- Opens the file in LibreOffice Calc
"""
import os
import subprocess
import shlex
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_014'

# Install dependencies
subprocess.run(['pip3', 'install', 'openpyxl'], capture_output=True)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Loan Calculator"

# ── Styling helpers ──
header_font = Font(name="Calibri", size=12, bold=True)
value_font = Font(name="Calibri", size=11)
section_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
section_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ── Column widths ──
ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 16
ws.column_dimensions["E"].width = 18

# ── Input Section (rows 1-4) ──
# Row 1: Title
ws.merge_cells("A1:E1")
ws["A1"] = "Loan Calculator"
ws["A1"].font = Font(name="Calibri", size=14, bold=True)
ws["A1"].alignment = Alignment(horizontal="center")

# Row 2: Principal
ws["A2"] = "Principal"
ws["A2"].font = header_font
ws["B2"] = 250000
ws["B2"].font = value_font
ws["B2"].number_format = '$#,##0.00'

# Row 3: Annual Rate
ws["A3"] = "Annual Rate"
ws["A3"].font = header_font
ws["B3"] = 0.065
ws["B3"].font = value_font
ws["B3"].number_format = '0.00%'

# Row 4: Term (Years)
ws["A4"] = "Term (Years)"
ws["A4"].font = header_font
ws["B4"] = 30
ws["B4"].font = value_font
ws["B4"].number_format = '0'

# Row 5: Monthly Payment placeholder (label only, no formula)
ws["A5"] = "Monthly Payment"
ws["A5"].font = header_font
# B5 left empty - agent must fill with PMT formula

# Row 6: blank separator

# ── Amortization Table Headers (row 7) ──
headers = ["Payment #", "Payment", "Interest", "Principal", "Balance"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=7, column=c, value=h)
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

# Save
output_path = os.path.join(WORKDIR, f"{TASK_ID}.xlsx")
wb.save(output_path)
print(f"Saved initial file to {output_path}")

# Launch LibreOffice Calc with the file
def launch_gui(command, delay_sec=1.0):
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

launch_gui(f'libreoffice --calc "{output_path}"', delay_sec=2.0)
print("LibreOffice Calc launched.")
