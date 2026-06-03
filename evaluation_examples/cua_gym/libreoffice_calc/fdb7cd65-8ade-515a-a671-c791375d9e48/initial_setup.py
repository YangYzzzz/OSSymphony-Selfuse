"""
Initial Setup: Track professional development hours for teaching staff
Task ID: calc_edu_professional_dev_log_035
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_professional_dev_log_035'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: PDLog ---
    ws = wb.active
    ws.title = 'PDLog'

    # Headers in row 1
    headers = [
        'Teacher Name',      # A
        'Workshop Hours',    # B
        'Online Course Hours',  # C
        'Conference Hours',  # D
        'Total Hours',       # E
        'Remaining',         # F
        'Pct Complete',      # G
        'Status',            # H
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Required hours label and value
    ws['I1'] = 'Required Hours:'
    ws['I1'].font = Font(bold=True)
    ws['J1'] = 40

    # Column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 13
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 13
    ws.column_dimensions['H'].width = 13

    # 25 teachers with realistic PD hours (B, C, D filled; E, F, G, H empty)
    teachers = [
        ('Sarah Chen',         12, 10,  8),
        ('Marcus Johnson',      8, 15,  5),
        ('Emily Rodriguez',    16,  8, 12),
        ('David Patel',         4, 10,  6),
        ('Aisha Williams',     20,  8, 10),
        ('James O\'Brien',      6, 12,  4),
        ('Priya Sharma',       14, 14,  8),
        ('Michael Torres',     10,  6,  3),
        ('Linda Washington',   18, 12, 10),
        ('Robert Kim',          2,  8,  0),
        ('Natalie Gomez',      10, 10,  5),
        ('Christopher Lee',    15,  8, 10),
        ('Amanda Foster',       6,  6,  4),
        ('Daniel Nguyen',      20, 10, 12),
        ('Jessica Martinez',    8, 14,  6),
        ('Kevin Brown',        12,  6,  8),
        ('Michelle Davis',      4, 10,  2),
        ('Thomas Wilson',      16, 10, 12),
        ('Rachel Clark',        8,  8,  6),
        ('Anthony Robinson',    0,  6,  0),
        ('Stephanie Hall',     14, 12,  8),
        ('Brian Scott',        10,  8,  4),
        ('Nicole Adams',       18,  8, 14),
        ('Ryan Baker',          6, 12,  2),
        ('Megan Campbell',     12, 10, 10),
    ]

    for row_idx, (name, workshop, online, conference) in enumerate(teachers, 2):
        ws.cell(row=row_idx, column=1, value=name)
        ws.cell(row=row_idx, column=2, value=workshop)
        ws.cell(row=row_idx, column=3, value=online)
        ws.cell(row=row_idx, column=4, value=conference)
        # E, F, G, H intentionally left empty (task will fill these)

    # Row 26: totals row label (empty E, F, G, H for totals)
    ws.cell(row=26, column=1, value='TOTALS')
    ws.cell(row=26, column=1).font = Font(bold=True)

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
