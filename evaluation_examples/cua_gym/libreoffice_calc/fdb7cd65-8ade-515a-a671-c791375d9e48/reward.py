"""
Reward Script: Track professional development hours for teaching staff
Task ID: calc_edu_professional_dev_log_035
Domain: libreoffice_calc

Scoring:
  Component 1: Total Hours formulas in E2:E25 (=B+C+D pattern)         — 0.25 pts
  Component 2: Remaining Hours formulas in F2:F25 (=MAX(0,$J$1-E))     — 0.20 pts
  Component 3: Pct Complete formulas in G2:G25 + 0.00% format           — 0.20 pts
  Component 4: Status formulas in H2:H25 (=IF(E>=J1,"Complete","Incomplete")) — 0.20 pts
  Component 5: Totals row E26 = SUM(E2:E25)                             — 0.05 pts
  Component 6: Conditional formatting on E2:H26 red when Incomplete     — 0.10 pts
  Total: 1.00
"""

import os
import openpyxl
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_professional_dev_log_035'


def normalize_formula(f):
    """Normalize a formula string for comparison: uppercase, no spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'PDLog' not in wb.sheetnames:
        print("CRITICAL: Sheet 'PDLog' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['PDLog']

    # -------------------------------------------------------------------------
    # Component 1: Total Hours formulas in E2:E25 (0.25 points)
    # Each row n should have formula =Bn+Cn+Dn (or equivalent sum)
    # -------------------------------------------------------------------------
    try:
        e_formula_count = 0
        e_total_rows = 24  # rows 2-25
        for row in range(2, 26):
            cell_val = ws.cell(row=row, column=5).value  # column E
            if cell_val is not None and isinstance(cell_val, str):
                norm = normalize_formula(cell_val)
                # Accept =Bn+Cn+Dn or =Bn+Cn+Dn pattern with any permutation
                # Must reference B, C, D columns for the same row
                b_ref = f'B{row}'
                c_ref = f'C{row}'
                d_ref = f'D{row}'
                if (b_ref in norm and c_ref in norm and d_ref in norm and
                        '+' in norm):
                    e_formula_count += 1

        if e_formula_count == e_total_rows:
            print(f"PASS: Component 1 — Total Hours formulas (E2:E25): all {e_formula_count}/{e_total_rows} rows have correct =B+C+D formula (0.25 pts)")
            total_score += 0.25
        elif e_formula_count >= e_total_rows * 0.8:
            partial = round(0.25 * e_formula_count / e_total_rows, 4)
            print(f"PARTIAL: Component 1 — Total Hours formulas: {e_formula_count}/{e_total_rows} rows correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Total Hours formulas: only {e_formula_count}/{e_total_rows} rows have =B+C+D formula in column E")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Remaining Hours formulas in F2:F25 (0.20 points)
    # Each row n should have formula =MAX(0,$J$1-En) (uses absolute $J$1)
    # -------------------------------------------------------------------------
    try:
        f_formula_count = 0
        f_total_rows = 24  # rows 2-25
        for row in range(2, 26):
            cell_val = ws.cell(row=row, column=6).value  # column F
            if cell_val is not None and isinstance(cell_val, str):
                norm = normalize_formula(cell_val)
                e_ref = f'E{row}'
                # Must contain MAX, reference $J$1 (absolute), and E row reference
                if ('MAX' in norm and '$J$1' in norm and e_ref in norm):
                    f_formula_count += 1

        if f_formula_count == f_total_rows:
            print(f"PASS: Component 2 — Remaining Hours formulas (F2:F25): all {f_formula_count}/{f_total_rows} rows have correct =MAX(0,$J$1-E) formula (0.20 pts)")
            total_score += 0.20
        elif f_formula_count >= f_total_rows * 0.8:
            partial = round(0.20 * f_formula_count / f_total_rows, 4)
            print(f"PARTIAL: Component 2 — Remaining Hours formulas: {f_formula_count}/{f_total_rows} rows correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Remaining Hours formulas: only {f_formula_count}/{f_total_rows} rows have =MAX(0,$J$1-E) formula in column F")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Pct Complete formulas + percentage format in G2:G25 (0.20 points)
    # Each row n should have formula =En/$J$1 with number_format including '%'
    # -------------------------------------------------------------------------
    try:
        g_formula_count = 0
        g_format_count = 0
        g_total_rows = 24  # rows 2-25
        for row in range(2, 26):
            cell = ws.cell(row=row, column=7)  # column G
            cell_val = cell.value
            if cell_val is not None and isinstance(cell_val, str):
                norm = normalize_formula(cell_val)
                e_ref = f'E{row}'
                # Must reference E row and $J$1
                if (e_ref in norm and '$J$1' in norm and '/' in norm):
                    g_formula_count += 1
            # Check percentage format
            if cell.number_format and '%' in cell.number_format:
                g_format_count += 1

        formula_ok = g_formula_count == g_total_rows
        format_ok = g_format_count >= g_total_rows * 0.8

        if formula_ok and format_ok:
            print(f"PASS: Component 3 — Pct Complete formulas (G2:G25): all {g_formula_count}/{g_total_rows} formulas correct, {g_format_count} cells with % format (0.20 pts)")
            total_score += 0.20
        elif formula_ok and not format_ok:
            print(f"PARTIAL: Component 3 — Pct Complete: formulas correct but only {g_format_count}/{g_total_rows} cells have % format (0.10 pts)")
            total_score += 0.10
        elif not formula_ok and g_formula_count >= g_total_rows * 0.8:
            partial = round(0.20 * g_formula_count / g_total_rows, 4)
            print(f"PARTIAL: Component 3 — Pct Complete formulas: {g_formula_count}/{g_total_rows} rows correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Pct Complete: only {g_formula_count}/{g_total_rows} formula rows and {g_format_count} % formatted cells")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: Status formulas in H2:H25 (0.20 points)
    # Each row n should have =IF(En>=$J$1,"Complete","Incomplete")
    # -------------------------------------------------------------------------
    try:
        h_formula_count = 0
        h_total_rows = 24  # rows 2-25
        for row in range(2, 26):
            cell_val = ws.cell(row=row, column=8).value  # column H
            if cell_val is not None and isinstance(cell_val, str):
                norm = normalize_formula(cell_val)
                e_ref = f'E{row}'
                # Must contain IF, reference E row and $J$1, and contain "COMPLETE"/"INCOMPLETE"
                if ('IF(' in norm and e_ref in norm and '$J$1' in norm and
                        'COMPLETE' in norm and 'INCOMPLETE' in norm):
                    h_formula_count += 1

        if h_formula_count == h_total_rows:
            print(f"PASS: Component 4 — Status formulas (H2:H25): all {h_formula_count}/{h_total_rows} rows have IF formula (0.20 pts)")
            total_score += 0.20
        elif h_formula_count >= h_total_rows * 0.8:
            partial = round(0.20 * h_formula_count / h_total_rows, 4)
            print(f"PARTIAL: Component 4 — Status formulas: {h_formula_count}/{h_total_rows} rows correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Status formulas: only {h_formula_count}/{h_total_rows} rows have correct IF formula in column H")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -------------------------------------------------------------------------
    # Component 5: Totals row E26 = SUM(E2:E25) (0.05 points)
    # -------------------------------------------------------------------------
    try:
        e26_val = ws.cell(row=26, column=5).value  # E26
        if e26_val is not None and isinstance(e26_val, str):
            norm = normalize_formula(e26_val)
            # Accept =SUM(E2:E25) or similar
            if 'SUM' in norm and 'E2' in norm and 'E25' in norm:
                print(f"PASS: Component 5 — Totals row E26 contains SUM formula: {repr(e26_val)} (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 5 — E26 formula '{e26_val}' does not match expected =SUM(E2:E25) pattern")
        else:
            print(f"FAIL: Component 5 — E26 is not a formula (value: {repr(e26_val)})")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # -------------------------------------------------------------------------
    # Component 6: Conditional formatting on E2:H26 with red background when H="Incomplete" (0.10 points)
    # -------------------------------------------------------------------------
    try:
        # Counts of matching sub-conditions found across all CF rules
        formula_match_count = 0   # number of rules referencing H col + "Incomplete"
        range_match_count = 0     # number of matching rules with correct E2:H26 range
        red_fill_match_count = 0  # number of matching rules with red fill

        for cf in ws.conditional_formatting:
            cf_str = str(cf)
            for rule in ws.conditional_formatting[cf]:
                rule_formulas = rule.formula if rule.formula else []
                for form in rule_formulas:
                    norm_form = normalize_formula(form)
                    if 'H' in norm_form and 'INCOMPLETE' in norm_form:
                        formula_match_count += 1
                        if 'E2' in cf_str and 'H26' in cf_str:
                            range_match_count += 1
                        if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                            try:
                                fill_color = rule.dxf.fill.fgColor.rgb
                                if fill_color and 'FF0000' in fill_color.upper():
                                    red_fill_match_count += 1
                            except Exception:
                                pass

        # Derive booleans from actual counts (not hardcoded)
        has_correct_formula = formula_match_count >= 1
        has_correct_range = range_match_count >= 1
        has_red_fill = red_fill_match_count >= 1

        if has_correct_formula and has_correct_range and has_red_fill:
            print(f"PASS: Component 6 — Conditional formatting: red background on E2:H26 when H='Incomplete' (0.10 pts)")
            total_score += 0.10
        elif has_correct_formula and has_red_fill:
            print(f"PARTIAL: Component 6 — Conditional formatting has correct formula and red fill, range differs from E2:H26 (0.05 pts)")
            total_score += 0.05
        elif has_correct_formula:
            print(f"PARTIAL: Component 6 — Conditional formatting formula references H/Incomplete but missing red fill (0.03 pts)")
            total_score += 0.03
        else:
            print(f"FAIL: Component 6 — No conditional formatting found that references H column with 'Incomplete' (formula_count={formula_match_count})")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
