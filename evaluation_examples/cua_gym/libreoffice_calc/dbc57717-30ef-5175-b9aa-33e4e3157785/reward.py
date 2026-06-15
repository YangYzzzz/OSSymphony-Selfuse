"""
Reward Script: Organize semester files into course folders, create index and summary
Task ID: osworld_multi_apps_doc_desktop_organize_011
Domain: multi_apps (os + libreoffice_calc + libreoffice_writer)
Scoring:
  Component 1: 4 course folders created on Desktop (0.25 pts)
  Component 2: All 24 files moved into correct course folders (0.35 pts)
  Component 3: semester_index.ods on Desktop with 24 data rows and required columns (0.25 pts)
  Component 4: semester_summary.odt on Desktop with H1 title and 4 H2 course sections (0.15 pts)
  Total: 1.0
"""

import os
import shutil

DESKTOP = '/home/user/Desktop'
TASK_ID = 'osworld_multi_apps_doc_desktop_organize_011'

EXPECTED_COURSES = ['CS101', 'MATH202', 'ENG301', 'HIST401']

# The 24 files from the initial state — map from course prefix to expected filenames
EXPECTED_FILES = {
    'CS101': [
        'CS101_Assignment1_Variables.odt',
        'CS101_Assignment2_Loops.odt',
        'CS101_FinalProject_Proposal.odt',
        'CS101_Lab3_Functions.odt',
        'CS101_Lecture_Notes_Week5.odp',
        'CS101_ProblemSet1.ods',
        'CS101_Reading_AlgorithmsIntro.pdf',
    ],
    'MATH202': [
        'MATH202_FinalExam_Practice.odt',
        'MATH202_Homework1_Integrals.odt',
        'MATH202_Homework2_Series.odt',
        'MATH202_MidtermReview.odp',
        'MATH202_ProblemSet_Week4.ods',
        'MATH202_Reading_TaylorSeries.pdf',
    ],
    'ENG301': [
        'ENG301_Essay1_TechReport.odt',
        'ENG301_Essay2_Analysis.odt',
        'ENG301_FinalPaper_Draft.odt',
        'ENG301_Presentation_Research.odp',
        'ENG301_Reading_StyleGuide.pdf',
    ],
    'HIST401': [
        'HIST401_Essay1_WWI_Causes.odt',
        'HIST401_Essay2_ColdWar.odt',
        'HIST401_FinalThesis_Outline.odt',
        'HIST401_Presentation_Timeline.odp',
        'HIST401_Reading_PrimarySource.pdf',
        'HIST401_ResearchNotes.ods',
    ],
}

TOTAL_FILES = sum(len(v) for v in EXPECTED_FILES.values())  # 24


def verify_task():
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # -----------------------------------------------------------------------
    # Component 1: 4 course folders created on Desktop (0.25 pts)
    # All 4 folders must exist as directories on the Desktop.
    # FAILS on initial (no folders) -> PASSES on golden (4 folders present)
    # -----------------------------------------------------------------------
    try:
        folders_found = []
        for course in EXPECTED_COURSES:
            folder_path = os.path.join(DESKTOP, course)
            if os.path.isdir(folder_path):
                folders_found.append(course)
        if len(folders_found) == len(EXPECTED_COURSES):
            print(f"PASS: Component 1 — all 4 course folders found: {folders_found} (0.25 pts)")
            total_score += 0.25
        else:
            missing = [c for c in EXPECTED_COURSES if c not in folders_found]
            print(f"FAIL: Component 1 — missing course folders: {missing} (found {len(folders_found)}/4)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: All 24 files moved to correct course folders (0.35 pts)
    # Each file should be present in its matching course subfolder.
    # FAILS on initial (files on Desktop root) -> PASSES on golden (files in folders)
    # Partial: at least 18/24 files correctly placed earns 0.20 pts;
    #          all 24/24 earns full 0.35 pts.
    # -----------------------------------------------------------------------
    try:
        files_correct = 0
        files_total = TOTAL_FILES
        misplaced = []
        for course, filenames in EXPECTED_FILES.items():
            for fname in filenames:
                expected_path = os.path.join(DESKTOP, course, fname)
                # Also verify that the file is NOT still sitting on the Desktop root
                root_path = os.path.join(DESKTOP, fname)
                in_folder = os.path.isfile(expected_path)
                still_on_desktop = os.path.isfile(root_path)
                if in_folder and not still_on_desktop:
                    files_correct += 1
                else:
                    misplaced.append(f"{fname} (in_folder={in_folder}, on_desktop={still_on_desktop})")
        if files_correct == files_total:
            print(f"PASS: Component 2 — all {files_total} files moved to correct course folders (0.35 pts)")
            total_score += 0.35
        elif files_correct >= 18:
            print(f"PARTIAL: Component 2 — {files_correct}/{files_total} files correctly placed (0.20 pts)")
            print(f"  Misplaced: {misplaced[:5]}")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 — only {files_correct}/{files_total} files correctly placed (0 pts)")
            print(f"  Misplaced sample: {misplaced[:5]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: semester_index.ods on Desktop with 24 data rows and correct columns (0.25 pts)
    # The file must exist AND contain headers: Filename, Course, File_Type, Moved_To
    # AND have exactly 24 data rows (1 per file).
    # FAILS on initial (file doesn't exist) -> PASSES on golden
    # -----------------------------------------------------------------------
    try:
        index_path = os.path.join(DESKTOP, 'semester_index.ods')
        if not os.path.isfile(index_path):
            print(f"FAIL: Component 3 — semester_index.ods not found at {index_path}")
        else:
            # The file is actually stored in xlsx format despite the .ods extension
            # Copy to a temp .xlsx path to load with openpyxl
            tmp_path = '/tmp/semester_index_check.xlsx'
            shutil.copy(index_path, tmp_path)
            try:
                import openpyxl
                wb = openpyxl.load_workbook(tmp_path)
                ws = wb.active
                headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
                required_headers = {'Filename', 'Course', 'File_Type', 'Moved_To'}
                # Normalize headers: case-insensitive check
                headers_set = {str(h).strip() if h else '' for h in headers}
                headers_match = required_headers.issubset(headers_set)
                # Count data rows (rows 2..max_row with non-empty Filename)
                data_rows = 0
                for r in range(2, ws.max_row + 1):
                    fname_val = ws.cell(r, 1).value
                    if fname_val is not None and str(fname_val).strip():
                        data_rows += 1

                if headers_match and data_rows == 24:
                    print(f"PASS: Component 3 — semester_index.ods found with correct headers and {data_rows} data rows (0.25 pts)")
                    total_score += 0.25
                elif headers_match and data_rows >= 20:
                    print(f"PARTIAL: Component 3 — semester_index.ods headers OK but {data_rows}/24 data rows (0.15 pts)")
                    total_score += 0.15
                else:
                    print(f"FAIL: Component 3 — semester_index.ods: headers_match={headers_match}, data_rows={data_rows}")
                    print(f"  Headers found: {headers}")
            except Exception as inner_e:
                # Might be a true ODS file; try XML parsing fallback
                try:
                    import zipfile, xml.etree.ElementTree as ET
                    with zipfile.ZipFile(index_path, 'r') as z:
                        content = z.read('content.xml').decode('utf-8')
                    root = ET.fromstring(content)
                    ns_table = 'urn:oasis:names:tc:opendocument:xmlns:table:1.0'
                    ns_text = 'urn:oasis:names:tc:opendocument:xmlns:text:1.0'
                    ns_office = 'urn:oasis:names:tc:opendocument:xmlns:office:1.0'
                    ns = {'table': ns_table, 'text': ns_text, 'office': ns_office}
                    rows = root.findall(f'.//{{{ns_table}}}table-row')
                    data_rows = max(0, len(rows) - 1)  # subtract header row
                    if data_rows >= 20:
                        print(f"PARTIAL: Component 3 — ODS semester_index detected with ~{data_rows} rows (0.15 pts)")
                        total_score += 0.15
                    else:
                        print(f"FAIL: Component 3 — ODS semester_index found {data_rows} rows (need 24)")
                except Exception as xml_e:
                    print(f"FAIL: Component 3 — could not parse semester_index.ods: xlsx_err={inner_e}, xml_err={xml_e}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: semester_summary.odt on Desktop with H1 title and 4 H2 sections (0.15 pts)
    # Must have H1 containing 'Semester File Organization' and 4 H2 sections for each course.
    # FAILS on initial (file doesn't exist) -> PASSES on golden
    # -----------------------------------------------------------------------
    try:
        summary_path = os.path.join(DESKTOP, 'semester_summary.odt')
        if not os.path.isfile(summary_path):
            print(f"FAIL: Component 4 — semester_summary.odt not found at {summary_path}")
        else:
            try:
                from docx import Document
                doc = Document(summary_path)
                headings = [(p.style.name, p.text) for p in doc.paragraphs
                            if p.style.name.startswith('Heading')]
                # Check H1 title
                h1_texts = [text for style, text in headings if '1' in style]
                h2_texts = [text for style, text in headings if '2' in style]
                has_h1_title = any('semester file organization' in t.lower() for t in h1_texts)
                # Check that all 4 courses appear in H2 sections
                h2_all_text = ' '.join(h2_texts).upper()
                courses_in_h2 = sum(1 for c in EXPECTED_COURSES if c in h2_all_text)

                if has_h1_title and courses_in_h2 == 4:
                    print(f"PASS: Component 4 — semester_summary.odt H1='{h1_texts}' and {courses_in_h2}/4 courses in H2 (0.15 pts)")
                    total_score += 0.15
                elif courses_in_h2 >= 2:
                    print(f"PARTIAL: Component 4 — semester_summary.odt has_h1={has_h1_title}, courses_in_h2={courses_in_h2}/4 (0.08 pts)")
                    total_score += 0.08
                else:
                    print(f"FAIL: Component 4 — semester_summary.odt: has_h1_title={has_h1_title}, h2_courses={courses_in_h2}/4")
                    print(f"  H2 texts: {h2_texts}")
            except Exception as inner_e:
                print(f"FAIL: Component 4 — could not parse semester_summary.odt: {inner_e}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


verify_task()
