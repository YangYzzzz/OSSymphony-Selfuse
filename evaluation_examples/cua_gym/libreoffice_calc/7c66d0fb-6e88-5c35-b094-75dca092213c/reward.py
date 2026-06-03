"""
Reward Script: SSN Re-padding and Formatting
Task ID: calc_gen_data_cleanup_059
Domain: libreoffice_calc

Task: Import HR data with numeric SSNs (some 7-9 digits), re-pad to 9 digits, apply
      SSN display format in column D (XXX-XX-XXXX) and masked format in column E (***-**-XXXX).

Scoring:
  Component 1 (0.6): D2:D101 all contain SSN formatted formula using LEFT/TEXT/MID/RIGHT
  Component 2 (0.4): E2:E101 all contain SSN masked formula with '***-**-' prefix
  Total: 1.0

  Data integrity (columns A/B/C/F intact) is used as a precondition gate, not a scoring component,
  because these are pre-existing conditions — the initial file already has them.
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_data_cleanup_059'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that HRData sheet exists (precondition gate)
    if 'HRData' not in wb.sheetnames:
        print("CRITICAL: 'HRData' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['HRData']

    # Verify the file has data rows (precondition gate)
    if ws.max_row < 2:
        print("CRITICAL: HRData sheet has no data rows")
        print("REWARD: 0.0")
        return 0.0

    # Data integrity gate: check that columns A, B, C, F are not corrupted
    # This is a precondition check (not scored), because these columns exist in the initial file
    try:
        header_a = ws.cell(row=1, column=1).value
        header_f = ws.cell(row=1, column=6).value
        if header_a != 'Emp ID' or header_f != 'Department':
            print(f"CRITICAL: Data integrity failed — headers changed: A={repr(header_a)}, F={repr(header_f)}")
            print("REWARD: 0.0")
            return 0.0
    except Exception as e:
        print(f"CRITICAL: Cannot verify data integrity: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: D2:D101 contains SSN formatted formula for all 100 rows (0.6 points)
    # Formula must use TEXT to zero-pad to 9 digits and format as XXX-XX-XXXX
    # Expected pattern: =LEFT(TEXT(Cn,"000000000"),3)&"-"&MID(TEXT(Cn,"000000000"),4,2)&"-"&RIGHT(TEXT(Cn,"000000000"),4)
    try:
        d_formula_count = 0
        d_wrong_rows = []
        for row in range(2, 102):
            d_val = ws.cell(row=row, column=4).value
            if d_val is None:
                d_wrong_rows.append(f"D{row}: None")
                continue
            if not isinstance(d_val, str):
                d_wrong_rows.append(f"D{row}: not a string ({repr(d_val)})")
                continue

            # Normalize for comparison
            d_upper = d_val.upper().replace(" ", "")

            # Must reference TEXT(Cn, "000000000") for 9-digit padding
            has_text_pad = (
                f'TEXT(C{row},"000000000")' in d_val or
                f'TEXT(C{row},\'000000000\')' in d_val
            )

            # Must contain LEFT, MID, RIGHT for XXX-XX-XXXX formatting
            has_left = 'LEFT(' in d_upper
            has_mid = 'MID(' in d_upper
            has_right = 'RIGHT(' in d_upper

            # Must contain dashes as separators
            has_dashes = '"-"' in d_val or "'-'" in d_val

            if has_text_pad and has_left and has_mid and has_right and has_dashes:
                d_formula_count += 1
            else:
                d_wrong_rows.append(f"D{row}: formula missing required components: {repr(d_val[:60])}")

        if d_formula_count == 100:
            print(f"PASS: Component 1 — All 100 D column SSN format formulas present (0.6 pts)")
            total_score += 0.6
        elif d_formula_count > 0:
            partial = round(0.6 * d_formula_count / 100, 4)
            print(f"PARTIAL: Component 1 — {d_formula_count}/100 D column SSN format formulas present ({partial} pts)")
            if d_wrong_rows:
                print(f"  First wrong rows: {d_wrong_rows[:3]}")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — D column has no SSN format formulas (0.0 pts)")
            if d_wrong_rows:
                print(f"  First wrong rows: {d_wrong_rows[:3]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: E2:E101 contains SSN masked formula for all 100 rows (0.4 points)
    # Expected pattern: ="***-**-"&RIGHT(TEXT(Cn,"000000000"),4)
    try:
        e_formula_count = 0
        e_wrong_rows = []
        for row in range(2, 102):
            e_val = ws.cell(row=row, column=5).value
            if e_val is None:
                e_wrong_rows.append(f"E{row}: None")
                continue
            if not isinstance(e_val, str):
                e_wrong_rows.append(f"E{row}: not a string ({repr(e_val)})")
                continue

            # Normalize for comparison
            e_upper = e_val.upper().replace(" ", "")

            # Must reference TEXT(Cn, "000000000") for 9-digit padding
            has_text_pad = (
                f'TEXT(C{row},"000000000")' in e_val or
                f'TEXT(C{row},\'000000000\')' in e_val
            )

            # Must contain the ***-**- mask prefix
            has_mask = '***-**-' in e_val

            # Must contain RIGHT for the last 4 digits
            has_right = 'RIGHT(' in e_upper

            if has_text_pad and has_mask and has_right:
                e_formula_count += 1
            else:
                e_wrong_rows.append(f"E{row}: formula missing required components: {repr(e_val[:60])}")

        if e_formula_count == 100:
            print(f"PASS: Component 2 — All 100 E column SSN masked formulas present (0.4 pts)")
            total_score += 0.4
        elif e_formula_count > 0:
            partial = round(0.4 * e_formula_count / 100, 4)
            print(f"PARTIAL: Component 2 — {e_formula_count}/100 E column SSN masked formulas present ({partial} pts)")
            if e_wrong_rows:
                print(f"  First wrong rows: {e_wrong_rows[:3]}")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — E column has no SSN masked formulas (0.0 pts)")
            if e_wrong_rows:
                print(f"  First wrong rows: {e_wrong_rows[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
