"""
Initial Setup: HR data with SSN raw values needing cleanup
Task ID: calc_gen_data_cleanup_059
Domain: libreoffice_calc

Creates an HRData spreadsheet with 100 employee records where:
- Column C (SSN Raw) contains numeric SSN values that have lost leading zeros
  (some appear as 7 or 8 digits instead of 9)
- Columns D and E (SSN Formatted and SSN Masked) are empty
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_data_cleanup_059'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'HRData'

    # Headers
    headers = ['Emp ID', 'Name', 'SSN Raw', 'SSN Formatted', 'SSN Masked', 'Department']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic employee data
    # SSN Raw values: some with leading zeros stripped (7-8 digit numbers),
    # some full 9-digit numbers. Stored as integers (simulating import truncation).
    employees = [
        ('EMP001', 'Sarah Chen',          23456789,  '', '', 'Engineering'),
        ('EMP002', 'Marcus Johnson',      987654321, '', '', 'Marketing'),
        ('EMP003', 'Priya Patel',         1234567,   '', '', 'Finance'),
        ('EMP004', 'James O\'Brien',      456789012, '', '', 'Engineering'),
        ('EMP005', 'Aisha Williams',      78901234,  '', '', 'HR'),
        ('EMP006', 'Roberto Garcia',      345678901, '', '', 'Sales'),
        ('EMP007', 'Linda Nguyen',        9012345,   '', '', 'Operations'),
        ('EMP008', 'David Kim',           890123456, '', '', 'Engineering'),
        ('EMP009', 'Rachel Thompson',     67890123,  '', '', 'Marketing'),
        ('EMP010', 'Ahmed Hassan',        234567890, '', '', 'Finance'),
        ('EMP011', 'Sofia Rossi',         5678901,   '', '', 'HR'),
        ('EMP012', 'Tyler Brooks',        901234567, '', '', 'Sales'),
        ('EMP013', 'Mei-Ling Wu',         12345678,  '', '', 'Engineering'),
        ('EMP014', 'Carlos Mendez',       789012345, '', '', 'Operations'),
        ('EMP015', 'Anna Kowalski',       45678901,  '', '', 'Finance'),
        ('EMP016', 'Benjamin Okafor',     678901234, '', '', 'Marketing'),
        ('EMP017', 'Jessica Park',        3456789,   '', '', 'Engineering'),
        ('EMP018', 'Nathan Schultz',      567890123, '', '', 'HR'),
        ('EMP019', 'Fatima Al-Rashid',    89012345,  '', '', 'Sales'),
        ('EMP020', 'Michael Turner',      456789012, '', '', 'Finance'),
        ('EMP021', 'Yuki Tanaka',         7890123,   '', '', 'Operations'),
        ('EMP022', 'Diana Foster',        890123456, '', '', 'Engineering'),
        ('EMP023', 'Kevin Okonkwo',       34567890,  '', '', 'Marketing'),
        ('EMP024', 'Laura Fernandez',     123456789, '', '', 'HR'),
        ('EMP025', 'Christopher Lee',     901234567, '', '', 'Sales'),
        ('EMP026', 'Amara Diallo',        5678901,   '', '', 'Finance'),
        ('EMP027', 'Patrick Sullivan',    678901234, '', '', 'Engineering'),
        ('EMP028', 'Nadia Volkov',        23456789,  '', '', 'Operations'),
        ('EMP029', 'Emmanuel Adeyemi',    789012345, '', '', 'Marketing'),
        ('EMP030', 'Stephanie Mitchell',  12345678,  '', '', 'HR'),
        ('EMP031', 'Ivan Petrov',         456789012, '', '', 'Sales'),
        ('EMP032', 'Grace Osei',          8901234,   '', '', 'Finance'),
        ('EMP033', 'Andrew Crawford',     345678901, '', '', 'Engineering'),
        ('EMP034', 'Yolanda Reyes',       90123456,  '', '', 'Operations'),
        ('EMP035', 'Daniel Chukwu',       567890123, '', '', 'Marketing'),
        ('EMP036', 'Hannah Bergstrom',    4567890,   '', '', 'HR'),
        ('EMP037', 'Omar Abdullah',       901234567, '', '', 'Sales'),
        ('EMP038', 'Cynthia Nakamura',    23456789,  '', '', 'Finance'),
        ('EMP039', 'Frederick Boateng',   789012345, '', '', 'Engineering'),
        ('EMP040', 'Isabelle Dupont',     56789012,  '', '', 'Operations'),
        ('EMP041', 'Raymond Tran',        234567890, '', '', 'Marketing'),
        ('EMP042', 'Monica Ivanova',      6789012,   '', '', 'HR'),
        ('EMP043', 'Samuel Achebe',       890123456, '', '', 'Sales'),
        ('EMP044', 'Theresa Blackwood',   34567890,  '', '', 'Finance'),
        ('EMP045', 'Vladimir Sokolov',    123456789, '', '', 'Engineering'),
        ('EMP046', 'Oluwaseun Adeleke',   78901234,  '', '', 'Operations'),
        ('EMP047', 'Alexandra Romanov',   567890123, '', '', 'Marketing'),
        ('EMP048', 'Kwame Asante',        3456789,   '', '', 'HR'),
        ('EMP049', 'Elaine Hoffmann',     901234567, '', '', 'Sales'),
        ('EMP050', 'Marcus Webb',         45678901,  '', '', 'Finance'),
        ('EMP051', 'Chioma Eze',          678901234, '', '', 'Engineering'),
        ('EMP052', 'Thomas Lindgren',     8901234,   '', '', 'Operations'),
        ('EMP053', 'Valentina Cruz',      345678901, '', '', 'Marketing'),
        ('EMP054', 'Obinna Nwachukwu',    23456789,  '', '', 'HR'),
        ('EMP055', 'Sandra Eriksson',     789012345, '', '', 'Sales'),
        ('EMP056', 'Jerome Baptiste',     56789012,  '', '', 'Finance'),
        ('EMP057', 'Mei Zhao',            234567890, '', '', 'Engineering'),
        ('EMP058', 'Francis Mensah',      6789012,   '', '', 'Operations'),
        ('EMP059', 'Brigitte Moreau',     890123456, '', '', 'Marketing'),
        ('EMP060', 'Olumide Adewale',     12345678,  '', '', 'HR'),
        ('EMP061', 'Ingrid Holm',         567890123, '', '', 'Sales'),
        ('EMP062', 'Alejandro Vargas',    4567890,   '', '', 'Finance'),
        ('EMP063', 'Ngozi Okeke',         901234567, '', '', 'Engineering'),
        ('EMP064', 'Stefan Muller',       78901234,  '', '', 'Operations'),
        ('EMP065', 'Chiamaka Obi',        345678901, '', '', 'Marketing'),
        ('EMP066', 'Hiroshi Watanabe',    9012345,   '', '', 'HR'),
        ('EMP067', 'Portia Dlamini',      678901234, '', '', 'Sales'),
        ('EMP068', 'Gonzalo Ibarra',      23456789,  '', '', 'Finance'),
        ('EMP069', 'Ekaterina Smirnova',  789012345, '', '', 'Engineering'),
        ('EMP070', 'Babatunde Olawale',   56789012,  '', '', 'Operations'),
        ('EMP071', 'Simone Leclerc',      234567890, '', '', 'Marketing'),
        ('EMP072', 'Adeola Bakare',       7890123,   '', '', 'HR'),
        ('EMP073', 'Pieter van den Berg', 890123456, '', '', 'Sales'),
        ('EMP074', 'Felicia Asante',      12345678,  '', '', 'Finance'),
        ('EMP075', 'Raul Dominguez',      567890123, '', '', 'Engineering'),
        ('EMP076', 'Tomoko Ishida',       3456789,   '', '', 'Operations'),
        ('EMP077', 'Nnamdi Okereke',      901234567, '', '', 'Marketing'),
        ('EMP078', 'Astrid Johansson',    45678901,  '', '', 'HR'),
        ('EMP079', 'Mamadou Diagne',      678901234, '', '', 'Sales'),
        ('EMP080', 'Ludmila Novakova',    8901234,   '', '', 'Finance'),
        ('EMP081', 'Emeka Onyekachi',     345678901, '', '', 'Engineering'),
        ('EMP082', 'Cecilia Gutierrez',   23456789,  '', '', 'Operations'),
        ('EMP083', 'Takeshi Yamamoto',    789012345, '', '', 'Marketing'),
        ('EMP084', 'Adaeze Nwosu',        56789012,  '', '', 'HR'),
        ('EMP085', 'Markus Bauer',        234567890, '', '', 'Sales'),
        ('EMP086', 'Taiwo Adesanya',      6789012,   '', '', 'Finance'),
        ('EMP087', 'Renata Kowalczyk',    890123456, '', '', 'Engineering'),
        ('EMP088', 'Chukwuemeka Agu',     34567890,  '', '', 'Operations'),
        ('EMP089', 'Beatriz Oliveira',    123456789, '', '', 'Marketing'),
        ('EMP090', 'Segun Adeyinka',      78901234,  '', '', 'HR'),
        ('EMP091', 'Marlene Hauser',      567890123, '', '', 'Sales'),
        ('EMP092', 'Ifeanyi Okafor',      4567890,   '', '', 'Finance'),
        ('EMP093', 'Akosua Mensah',       901234567, '', '', 'Engineering'),
        ('EMP094', 'Dmitri Volkov',       23456789,  '', '', 'Operations'),
        ('EMP095', 'Zainab Musa',         789012345, '', '', 'Marketing'),
        ('EMP096', 'Bart Vermeer',        56789012,  '', '', 'HR'),
        ('EMP097', 'Kehinde Oladele',     234567890, '', '', 'Sales'),
        ('EMP098', 'Monika Novak',        7890123,   '', '', 'Finance'),
        ('EMP099', 'Ebuka Chukwudi',      890123456, '', '', 'Engineering'),
        ('EMP100', 'Anastasia Petrov',    12345678,  '', '', 'Operations'),
    ]

    for r, (emp_id, name, ssn_raw, ssn_fmt, ssn_mask, dept) in enumerate(employees, 2):
        ws.cell(row=r, column=1, value=emp_id)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=ssn_raw)   # numeric SSN (leading zeros stripped)
        # Columns D (4) and E (5) intentionally left empty — task is to populate them
        ws.cell(row=r, column=4, value=None)
        ws.cell(row=r, column=5, value=None)
        ws.cell(row=r, column=6, value=dept)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: HRData')
    print(f'  Rows: 101 (1 header + 100 data rows)')
    print(f'  Columns: Emp ID, Name, SSN Raw, SSN Formatted (empty), SSN Masked (empty), Department')


create_initial()
