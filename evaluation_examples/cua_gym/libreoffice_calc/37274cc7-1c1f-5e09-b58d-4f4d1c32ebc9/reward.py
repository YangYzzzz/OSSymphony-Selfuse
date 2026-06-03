"""
Reward Script: Apply decimal validation to cell C2 that only allows negative values
Task ID: calc_nrv_079
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): Data validation exists targeting C2
  Component 2 (0.4): Validation type is 'decimal' and operator is 'lessThan'
  Component 3 (0.3): Validation formula1 (threshold) is 0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_079'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Precondition: headers should be intact
    if ws['A1'].value != 'Date' or ws['B1'].value != 'Description' or ws['C1'].value != 'Debit':
        print(f"WARNING: Headers may have changed. A1={ws['A1'].value}, B1={ws['B1'].value}, C1={ws['C1'].value}")

    # Find data validation targeting C2
    target_dv = None
    validations = ws.data_validations.dataValidation
    for dv in validations:
        sqref_str = str(dv.sqref)
        if 'C2' in sqref_str:
            target_dv = dv
            break

    # Component 1: Data validation exists on C2 (0.3 points)
    try:
        if target_dv is not None:
            print(f"PASS: Component 1 — Data validation found targeting C2 (sqref: {target_dv.sqref}) (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — No data validation found targeting C2. Found {len(validations)} validation(s) total.")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Validation type is 'decimal' and operator is 'lessThan' (0.4 points)
    try:
        if target_dv is not None:
            dv_type = target_dv.type
            dv_operator = target_dv.operator
            type_ok = (dv_type == 'decimal')
            operator_ok = (dv_operator == 'lessThan')
            if type_ok and operator_ok:
                print(f"PASS: Component 2 — type='{dv_type}', operator='{dv_operator}' (0.4 pts)")
                total_score += 0.4
            else:
                print(f"FAIL: Component 2 — Expected type='decimal' and operator='lessThan', found type='{dv_type}', operator='{dv_operator}'")
        else:
            print(f"FAIL: Component 2 — No validation to check (depends on Component 1)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Validation formula1 (threshold) is 0 (0.3 points)
    try:
        if target_dv is not None:
            formula1_val = target_dv.formula1
            # formula1 might be a string "0" or numeric 0
            try:
                threshold = float(str(formula1_val))
                threshold_ok = (threshold == 0.0)
            except (ValueError, TypeError):
                threshold_ok = False
                threshold = formula1_val

            if threshold_ok:
                print(f"PASS: Component 3 — formula1 threshold = {formula1_val} (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 3 — Expected formula1=0, found '{formula1_val}'")
        else:
            print(f"FAIL: Component 3 — No validation to check (depends on Component 1)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
