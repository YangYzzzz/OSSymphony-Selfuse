"""
Initial Setup: Apply decimal validation to cell C2 for debit transactions
Task ID: calc_nrv_079
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_079'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Accounting ---
    ws = wb.active
    ws.title = 'Accounting'

    # Headers
    headers = ['Date', 'Description', 'Debit']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Data rows - realistic accounting debit entries
    # C2 is intentionally left EMPTY (task requires adding validation there)
    data = [
        ['2025-03-01', 'Office supplies purchase', None],       # Row 2 - C2 empty
        ['2025-03-03', 'Client lunch meeting', -87.50],         # Row 3
        ['2025-03-05', 'Software subscription renewal', -249.99],  # Row 4
        ['2025-03-07', 'Travel expenses - airport taxi', -42.00],  # Row 5
        ['2025-03-10', 'Equipment maintenance fee', -315.00],      # Row 6
        ['2025-03-12', 'Marketing materials printing', -178.25],   # Row 7
        ['2025-03-15', 'Employee training workshop', -450.00],     # Row 8
        ['2025-03-18', 'Courier service charges', -23.75],         # Row 9
        ['2025-03-20', 'Conference room rental', -600.00],         # Row 10
        ['2025-03-22', 'IT support contract payment', -1250.00],   # Row 11
        ['2025-03-25', 'Stationery bulk order', -95.40],           # Row 12
        ['2025-03-28', 'Building security deposit', -2000.00],     # Row 13
    ]

    date_format = 'yyyy-mm-dd'
    currency_format = '#,##0.00'

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0]).number_format = date_format
        ws.cell(row=r, column=2, value=row_data[1])
        cell_c = ws.cell(row=r, column=3, value=row_data[2])
        if row_data[2] is not None:
            cell_c.number_format = currency_format

    # Column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 16

    # NO data validation on C2 - that is what the task asks the agent to do

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
