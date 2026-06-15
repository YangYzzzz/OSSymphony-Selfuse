"""
Reward Script: Fix pivot table data field from SUM of Quantity to SUM of TotalPrice
Task ID: calc_pivot_079
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): Header B3 changed to "SUM of TotalPrice"
  Component 2 (0.5): Product TotalPrice values correct (B4-B8)
  Component 3 (0.2): Grand Total B9 = 125000
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_079'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Report sheet must exist
    if 'Report' not in wb.sheetnames:
        print("CRITICAL: 'Report' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Report']

    # Component 1: Header B3 changed to "SUM of TotalPrice" (0.3 points)
    # Initial has "SUM of Quantity" — this checks the actual change
    try:
        header_val = ws['B3'].value
        if header_val is not None and 'totalprice' in str(header_val).lower():
            print(f"PASS: Component 1 — B3 header contains 'TotalPrice': '{header_val}' (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — Expected header with 'TotalPrice', found: '{header_val}'")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Product TotalPrice values are correct in B4:B8 (0.5 points)
    # Initial has quantity sums (288, 319, 285, 295, 285) — golden has dollar values
    # Expected: ProductA=25000, ProductB=18000, ProductC=32000, ProductD=22000, ProductE=28000
    expected_values = {
        'B4': 25000,  # ProductA
        'B5': 18000,  # ProductB
        'B6': 32000,  # ProductC
        'B7': 22000,  # ProductD
        'B8': 28000,  # ProductE
    }
    try:
        correct_count = 0
        for cell_ref, expected_val in expected_values.items():
            actual_val = ws[cell_ref].value
            if actual_val is not None:
                try:
                    if abs(float(actual_val) - expected_val) < 1.0:
                        correct_count += 1
                    else:
                        print(f"  INFO: {cell_ref} expected {expected_val}, found {actual_val}")
                except (ValueError, TypeError):
                    print(f"  INFO: {cell_ref} not numeric: {actual_val}")
            else:
                print(f"  INFO: {cell_ref} is None")

        if correct_count == 5:
            print(f"PASS: Component 2 — All 5 product TotalPrice values correct (0.5 pts)")
            total_score += 0.5
        elif correct_count >= 3:
            partial = round(0.5 * correct_count / 5, 2)
            print(f"PARTIAL: Component 2 — {correct_count}/5 values correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {correct_count}/5 values correct")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Grand Total B9 = 125000 (0.2 points)
    # Initial has 1472 (sum of quantities) — golden has 125000 (sum of TotalPrice)
    try:
        grand_total = ws['B9'].value
        if grand_total is not None:
            try:
                if abs(float(grand_total) - 125000) < 1.0:
                    print(f"PASS: Component 3 — Grand Total B9 = {grand_total} (0.2 pts)")
                    total_score += 0.2
                else:
                    print(f"FAIL: Component 3 — Expected Grand Total ~125000, found: {grand_total}")
            except (ValueError, TypeError):
                print(f"FAIL: Component 3 — Grand Total not numeric: {grand_total}")
        else:
            print(f"FAIL: Component 3 — B9 is None")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
