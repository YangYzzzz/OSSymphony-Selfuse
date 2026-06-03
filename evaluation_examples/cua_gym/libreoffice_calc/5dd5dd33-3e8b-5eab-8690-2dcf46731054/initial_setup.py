"""
Initial Setup: Pivot table on Report sheet showing SUM of Quantity (wrong field)
Task ID: calc_pivot_079
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_079'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()

    # --- Sheet 1: Orders ---
    ws_orders = wb.active
    ws_orders.title = 'Orders'

    headers = ['OrderID', 'Product', 'Quantity', 'UnitPrice', 'TotalPrice']
    header_font = Font(bold=True, size=11, name='Calibri')
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws_orders.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, name='Calibri', color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Product catalog with unit prices
    products = {
        'ProductA': [45.00, 50.00, 55.00],
        'ProductB': [30.00, 35.00, 40.00],
        'ProductC': [60.00, 65.00, 70.00],
        'ProductD': [25.00, 28.00, 32.00],
        'ProductE': [80.00, 85.00, 90.00],
    }

    # We need specific TotalPrice sums:
    # ProductA = 25000, ProductB = 18000, ProductC = 32000
    # ProductD + ProductE = 50000

    # Build orders with controlled totals
    orders = []
    order_id = 1001

    # ProductA: target 25000
    # 50 orders, avg total = 500 each
    product_a_prices = [45.00, 50.00, 55.00]
    remaining_a = 25000.0
    for i in range(29):
        qty = random.randint(5, 15)
        price = random.choice(product_a_prices)
        total = qty * price
        remaining_a -= total
        orders.append([order_id, 'ProductA', qty, price, total])
        order_id += 1
    # Last order for ProductA to hit exact target
    last_qty_a = 10
    last_price_a = round(remaining_a / last_qty_a, 2)
    orders.append([order_id, 'ProductA', last_qty_a, last_price_a, round(remaining_a, 2)])
    order_id += 1

    # ProductB: target 18000
    remaining_b = 18000.0
    for i in range(29):
        qty = random.randint(5, 15)
        price = random.choice([30.00, 35.00, 40.00])
        total = qty * price
        remaining_b -= total
        orders.append([order_id, 'ProductB', qty, price, total])
        order_id += 1
    last_qty_b = 10
    last_price_b = round(remaining_b / last_qty_b, 2)
    orders.append([order_id, 'ProductB', last_qty_b, last_price_b, round(remaining_b, 2)])
    order_id += 1

    # ProductC: target 32000
    remaining_c = 32000.0
    for i in range(29):
        qty = random.randint(5, 15)
        price = random.choice([60.00, 65.00, 70.00])
        total = qty * price
        remaining_c -= total
        orders.append([order_id, 'ProductC', qty, price, total])
        order_id += 1
    last_qty_c = 10
    last_price_c = round(remaining_c / last_qty_c, 2)
    orders.append([order_id, 'ProductC', last_qty_c, last_price_c, round(remaining_c, 2)])
    order_id += 1

    # ProductD: target 22000
    remaining_d = 22000.0
    for i in range(29):
        qty = random.randint(5, 15)
        price = random.choice([25.00, 28.00, 32.00])
        total = qty * price
        remaining_d -= total
        orders.append([order_id, 'ProductD', qty, price, total])
        order_id += 1
    last_qty_d = 10
    last_price_d = round(remaining_d / last_qty_d, 2)
    orders.append([order_id, 'ProductD', last_qty_d, last_price_d, round(remaining_d, 2)])
    order_id += 1

    # ProductE: target 28000
    remaining_e = 28000.0
    for i in range(29):
        qty = random.randint(5, 15)
        price = random.choice([80.00, 85.00, 90.00])
        total = qty * price
        remaining_e -= total
        orders.append([order_id, 'ProductE', qty, price, total])
        order_id += 1
    last_qty_e = 10
    last_price_e = round(remaining_e / last_qty_e, 2)
    orders.append([order_id, 'ProductE', last_qty_e, last_price_e, round(remaining_e, 2)])
    order_id += 1

    # Shuffle orders for realism
    random.shuffle(orders)

    # Write orders data
    for r, row_data in enumerate(orders, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_orders.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c in (3,):
                cell.number_format = '0'
            elif c in (4, 5):
                cell.number_format = '$#,##0.00'

    # Column widths
    ws_orders.column_dimensions['A'].width = 12
    ws_orders.column_dimensions['B'].width = 14
    ws_orders.column_dimensions['C'].width = 12
    ws_orders.column_dimensions['D'].width = 12
    ws_orders.column_dimensions['E'].width = 14

    # Compute quantity sums per product for the pivot summary
    qty_sums = {}
    for row_data in orders:
        product = row_data[1]
        qty = row_data[2]
        qty_sums[product] = qty_sums.get(product, 0) + qty

    # --- Sheet 2: Report (Pivot Table - SUM of Quantity) ---
    ws_report = wb.create_sheet('Report')

    # Title
    ws_report.merge_cells('A1:C1')
    title_cell = ws_report['A1']
    title_cell.value = 'Pivot Table - Orders Summary'
    title_cell.font = Font(bold=True, size=14, name='Calibri')
    title_cell.alignment = Alignment(horizontal="center")

    # Pivot headers
    pivot_headers = ['Product', 'SUM of Quantity']
    for col, h in enumerate(pivot_headers, 1):
        cell = ws_report.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, size=11, name='Calibri', color="FFFFFF")
        cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Pivot data rows (sorted by product name)
    sorted_products = sorted(qty_sums.keys())
    row = 4
    total_qty = 0
    for product in sorted_products:
        ws_report.cell(row=row, column=1, value=product).border = thin_border
        cell_val = ws_report.cell(row=row, column=2, value=qty_sums[product])
        cell_val.border = thin_border
        cell_val.number_format = '#,##0'
        total_qty += qty_sums[product]
        row += 1

    # Grand Total row
    gt_label = ws_report.cell(row=row, column=1, value='Grand Total')
    gt_label.font = Font(bold=True)
    gt_label.border = thin_border
    gt_val = ws_report.cell(row=row, column=2, value=total_qty)
    gt_val.font = Font(bold=True)
    gt_val.border = thin_border
    gt_val.number_format = '#,##0'

    ws_report.column_dimensions['A'].width = 16
    ws_report.column_dimensions['B'].width = 20

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Orders: {len(orders)} rows')
    print(f'Quantity sums: {qty_sums}')
    print(f'Grand total quantity: {total_qty}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
