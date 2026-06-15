"""
Initial Setup: Build a summary label formula in Sheet2 referencing Q3 revenue from Sheet1
Task ID: osworld_calc_text_format_number_003
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_text_format_number_003'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Revenue Data ---
    ws1 = wb.active
    ws1.title = 'Sheet1'

    # Headers row
    headers = ['Quarter', 'Year', 'Region', 'Revenue', 'Notes']
    bold_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = bold_font

    # Quarterly revenue data — realistic business content
    # IMPORTANT: D2 must be the Q3 2024 total (as stated in task context)
    # Row 2 is Q3 2024, so Sheet1!D2 = 847235.50
    data = [
        ['Q3', 2024, 'North America',  847235.50, 'Record quarter'],  # row 2, D2 — referenced by task
        ['Q2', 2024, 'North America',  734195.20, 'Seasonal peak'],
        ['Q1', 2024, 'North America',  612480.75, 'Strong YoY growth'],
        ['Q4', 2023, 'North America',  755600.25, 'Year-end push'],
        ['Q3', 2023, 'North America',  702180.90, 'Steady performance'],
        ['Q2', 2023, 'North America',  668750.40, 'Mid-year growth'],
        ['Q1', 2023, 'North America',  521300.00, 'Baseline period'],
        ['Q4', 2022, 'North America',  689420.60, 'Strong finish'],
        ['Q3', 2022, 'North America',  643100.30, 'Growth quarter'],
        ['Q2', 2022, 'North America',  598750.80, 'Moderate gains'],
        ['Q1', 2022, 'North America',  552340.15, 'Slow start'],
        ['Q4', 2021, 'North America',  621890.45, 'Holiday season'],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 8
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 22

    # --- Sheet 2: Empty (task requires agent to add formula here) ---
    ws2 = wb.create_sheet('Sheet2')
    # Sheet2 must be completely empty — no formulas, no text
    # Agent must add formula to A1: ="Q3 2024 Revenue: $"&TEXT(Sheet1!D2,"#,##0.00")

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open LibreOffice Calc with the initial file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
