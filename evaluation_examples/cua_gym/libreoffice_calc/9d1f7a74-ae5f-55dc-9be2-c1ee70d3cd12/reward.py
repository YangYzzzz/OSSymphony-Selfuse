"""
Reward Script: Build a summary label formula in Sheet2 referencing Q3 2024 Revenue from Sheet1
Task ID: osworld_calc_text_format_number_003
Domain: libreoffice_calc
Scoring:
  Component 1: Sheet2 A1 contains a formula (not empty)                     — 0.3 points
  Component 2: Formula references Sheet1!D2 (the Q3 2024 revenue source)    — 0.3 points
  Component 3: Formula uses TEXT() with #,##0.00 format pattern              — 0.2 points
  Component 4: Formula includes 'Q3 2024 Revenue: $' label prefix           — 0.2 points
  Total: 1.0
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_text_format_number_003'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task: Sheet2 cell A1 must contain a formula that:
      - Is a formula (starts with '=')
      - References Sheet1!D2 (Q3 2024 Revenue total = 847235.5)
      - Uses TEXT() function with '#,##0.00' format (thousands sep + 2 decimal places)
      - Includes the label prefix 'Q3 2024 Revenue: $'
    """
    total_score = 0.0

    # Load workbook (formula mode to read formula strings)
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: Sheet2 must exist
    if 'Sheet2' not in wb.sheetnames:
        print("FAIL: Sheet2 does not exist in workbook")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws2 = wb['Sheet2']
    cell_a1 = ws2['A1']
    a1_value = cell_a1.value

    # Precondition gate: Sheet1 must exist and D2 must have data
    if 'Sheet1' not in wb.sheetnames:
        print("FAIL: Sheet1 does not exist in workbook")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws1 = wb['Sheet1']
    d2_value = ws1['D2'].value
    print(f"INFO: Sheet1!D2 value = {d2_value}")
    print(f"INFO: Sheet2!A1 raw value = {repr(a1_value)}")

    # Component 1: Sheet2 A1 contains a formula (starts with '=') (0.3 points)
    try:
        a1_is_formula = (
            a1_value is not None
            and isinstance(a1_value, str)
            and a1_value.strip().startswith('=')
        )
        if a1_is_formula:
            print(f"PASS: Component 1 — Sheet2 A1 contains a formula (0.3 pts)")
            total_score += 0.3
        else:
            if a1_value is None:
                print("FAIL: Component 1 — Sheet2 A1 is empty (expected a formula)")
            else:
                print(f"FAIL: Component 1 — Sheet2 A1 does not contain a formula; found: {repr(a1_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Formula references Sheet1!D2 (0.3 points)
    # The formula should pull Q3 2024 revenue from Sheet1!D2
    try:
        if a1_value and isinstance(a1_value, str):
            # Normalize the formula: remove spaces, uppercase for comparison
            formula_norm = a1_value.upper().replace(' ', '')
            # Accept various valid references: Sheet1!D2, Sheet1.D2
            refs_d2 = bool(
                re.search(r'SHEET1[!.]D2', formula_norm)
            )
            if refs_d2:
                print(f"PASS: Component 2 — Formula references Sheet1!D2 (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — Formula does not reference Sheet1!D2; formula: {repr(a1_value)}")
        else:
            print("FAIL: Component 2 — No formula to check for Sheet1!D2 reference")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Formula uses TEXT() with '#,##0.00' format pattern (0.2 points)
    # The task requires thousands separator AND 2 decimal places
    try:
        if a1_value and isinstance(a1_value, str):
            formula_norm = a1_value.upper().replace(' ', '')
            has_text_func = 'TEXT(' in formula_norm
            # Check for the required number format pattern: #,##0.00
            # Accept case-insensitive and with or without quotes
            has_format = bool(re.search(r'#,##0\.00', a1_value))
            if has_text_func and has_format:
                print(f"PASS: Component 3 — Formula uses TEXT() with '#,##0.00' format (0.2 pts)")
                total_score += 0.2
            elif has_text_func and not has_format:
                print(f"FAIL: Component 3 — Formula uses TEXT() but missing '#,##0.00' format pattern; found: {repr(a1_value)}")
            else:
                print(f"FAIL: Component 3 — Formula does not use TEXT() function; found: {repr(a1_value)}")
        else:
            print("FAIL: Component 3 — No formula to check for TEXT() usage")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Formula includes 'Q3 2024 Revenue: $' label prefix (0.2 points)
    # The output string should start with this specific label
    try:
        if a1_value and isinstance(a1_value, str):
            # Check that the formula string contains the label text as a literal
            # The label could be written in different casing or slight variations
            # Per task: exactly "Q3 2024 Revenue: $"
            has_q3_label = 'Q3 2024 Revenue: $' in a1_value
            if has_q3_label:
                print(f"PASS: Component 4 — Formula includes 'Q3 2024 Revenue: $' label prefix (0.2 pts)")
                total_score += 0.2
            else:
                # Try case-insensitive check for partial credit indication
                a1_upper = a1_value.upper()
                if 'Q3' in a1_upper and 'REVENUE' in a1_upper:
                    print(f"FAIL: Component 4 — Formula contains Q3/Revenue but not exact label 'Q3 2024 Revenue: $'; found: {repr(a1_value)}")
                else:
                    print(f"FAIL: Component 4 — Formula does not include 'Q3 2024 Revenue: $' label; found: {repr(a1_value)}")
        else:
            print("FAIL: Component 4 — No formula to check for label prefix")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point: run against the canonical task artifact on the VM
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
