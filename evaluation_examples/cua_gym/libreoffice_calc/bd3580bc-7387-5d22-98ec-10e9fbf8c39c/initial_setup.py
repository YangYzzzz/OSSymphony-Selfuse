"""
Initial Setup: Stockout cost analysis spreadsheet with raw data (no formulas)
Task ID: calc_ops_078
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_078'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Stockout ---
    ws = wb.active
    ws.title = 'Stockout'

    # Headers
    headers = [
        'SKU', 'Stockout Events', 'Avg Lost Units', 'Unit Margin',
        'Lost Sales Cost', 'Expedite Cost/Event', 'Total Expedite',
        'Goodwill Cost/Event', 'Total Goodwill', 'Total Stockout Cost'
    ]
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows - columns A, B, C, D, F, H have values; E, G, I, J are EMPTY (agent must add formulas)
    data = [
        # SKU, Stockout Events, Avg Lost Units, Unit Margin, (E empty), Expedite Cost/Event, (G empty), Goodwill Cost/Event, (I empty), (J empty)
        ['SKU-A', 3, 50, 15, None, 200, None, 100, None, None],
        ['SKU-B', 1, 200, 8, None, 500, None, 300, None, None],
        ['SKU-C', 5, 30, 25, None, 150, None, 75, None, None],
        ['SKU-D', 2, 100, 12, None, 350, None, 200, None, None],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            if val is not None:
                ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    col_widths = {'A': 10, 'B': 16, 'C': 16, 'D': 14, 'E': 16,
                  'F': 20, 'G': 16, 'H': 20, 'I': 16, 'J': 20}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Set row 1 height for wrapped headers
    ws.row_dimensions[1].height = 30

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
