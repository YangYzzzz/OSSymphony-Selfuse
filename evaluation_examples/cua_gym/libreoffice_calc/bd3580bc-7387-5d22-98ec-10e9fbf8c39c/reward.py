"""
Reward Script: Build a stockout cost analysis with formulas for lost sales,
expediting costs, goodwill loss, and total stockout cost.
Task ID: calc_ops_078
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Lost Sales Cost formulas (E2:E5) = B*C*D
  Component 2 (0.25): Total Expedite formulas (G2:G5) = B*F
  Component 3 (0.20): Total Goodwill formulas (I2:I5) = B*H
  Component 4 (0.25): Total Stockout Cost formulas (J2:J5) = E+G+I
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_078'

# Ground truth expected values (from task context)
EXPECTED_VALUES = {
    'E2': 2250, 'E3': 1600, 'E4': 3750, 'E5': 2400,
    'G2': 600,  'G3': 500,  'G4': 750,  'G5': 700,
    'I2': 300,  'I3': 300,  'I4': 375,  'I5': 400,
    'J2': 3150, 'J3': 2400, 'J4': 4875, 'J5': 3500,
}

# Expected formula patterns (normalized, uppercase, no spaces)
EXPECTED_FORMULAS = {
    'E': '=B{r}*C{r}*D{r}',
    'G': '=B{r}*F{r}',
    'I': '=B{r}*H{r}',
    'J': '=E{r}+G{r}+I{r}',
}


def normalize_formula(f):
    """Normalize formula for comparison: uppercase, no spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def check_cell_formula_or_value(ws, ws_data, coord, expected_formula, expected_value):
    """
    Check if a cell has the correct formula OR the correct computed value.
    Returns True if the formula string matches OR the cached value matches.
    """
    # Check formula string first
    cell_val = ws[coord].value
    if isinstance(cell_val, str) and cell_val.startswith('='):
        norm_actual = normalize_formula(cell_val)
        norm_expected = normalize_formula(expected_formula)
        if norm_actual == norm_expected:
            return True
        # Also accept equivalent formulations (e.g., different ordering)
        # For multiplication, order shouldn't matter for simple products

    # Check cached computed value (data_only mode)
    cached = ws_data[coord].value
    if cached is not None:
        try:
            if abs(float(cached) - expected_value) < 0.01:
                return True
        except (ValueError, TypeError):
            pass

    # Also check if the formula cell itself holds a numeric value
    # (file may have been saved with computed values)
    if cell_val is not None and not isinstance(cell_val, str):
        try:
            if abs(float(cell_val) - expected_value) < 0.01:
                return True
        except (ValueError, TypeError):
            pass

    return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Also load with data_only to get cached computed values
    try:
        wb_data = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"WARNING: Cannot load data_only: {e}")
        wb_data = wb

    # Check sheet exists
    if 'Stockout' not in wb.sheetnames:
        print("FAIL: 'Stockout' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Stockout']
    ws_data = wb_data['Stockout']

    # Component 1: Lost Sales Cost formulas E2:E5 = B*C*D (0.30 points)
    try:
        col_e_pass = 0
        for row in range(2, 6):
            coord = f'E{row}'
            expected_f = EXPECTED_FORMULAS['E'].format(r=row)
            expected_v = EXPECTED_VALUES[coord]
            if check_cell_formula_or_value(ws, ws_data, coord, expected_f, expected_v):
                col_e_pass += 1
                print(f"PASS: {coord} — formula/value correct (expected {expected_v})")
            else:
                actual = ws[coord].value
                print(f"FAIL: {coord} — expected formula '{expected_f}' or value {expected_v}, found: {actual}")
        comp1_score = 0.30 * (col_e_pass / 4)
        if col_e_pass > 0:
            total_score += comp1_score
            print(f"PASS: Component 1 — Lost Sales Cost: {col_e_pass}/4 cells ({comp1_score:.2f} pts)")
        else:
            print(f"FAIL: Component 1 — Lost Sales Cost: 0/4 cells correct")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Total Expedite formulas G2:G5 = B*F (0.25 points)
    try:
        col_g_pass = 0
        for row in range(2, 6):
            coord = f'G{row}'
            expected_f = EXPECTED_FORMULAS['G'].format(r=row)
            expected_v = EXPECTED_VALUES[coord]
            if check_cell_formula_or_value(ws, ws_data, coord, expected_f, expected_v):
                col_g_pass += 1
                print(f"PASS: {coord} — formula/value correct (expected {expected_v})")
            else:
                actual = ws[coord].value
                print(f"FAIL: {coord} — expected formula '{expected_f}' or value {expected_v}, found: {actual}")
        comp2_score = 0.25 * (col_g_pass / 4)
        if col_g_pass > 0:
            total_score += comp2_score
            print(f"PASS: Component 2 — Total Expedite: {col_g_pass}/4 cells ({comp2_score:.2f} pts)")
        else:
            print(f"FAIL: Component 2 — Total Expedite: 0/4 cells correct")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Total Goodwill formulas I2:I5 = B*H (0.20 points)
    try:
        col_i_pass = 0
        for row in range(2, 6):
            coord = f'I{row}'
            expected_f = EXPECTED_FORMULAS['I'].format(r=row)
            expected_v = EXPECTED_VALUES[coord]
            if check_cell_formula_or_value(ws, ws_data, coord, expected_f, expected_v):
                col_i_pass += 1
                print(f"PASS: {coord} — formula/value correct (expected {expected_v})")
            else:
                actual = ws[coord].value
                print(f"FAIL: {coord} — expected formula '{expected_f}' or value {expected_v}, found: {actual}")
        comp3_score = 0.20 * (col_i_pass / 4)
        if col_i_pass > 0:
            total_score += comp3_score
            print(f"PASS: Component 3 — Total Goodwill: {col_i_pass}/4 cells ({comp3_score:.2f} pts)")
        else:
            print(f"FAIL: Component 3 — Total Goodwill: 0/4 cells correct")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Total Stockout Cost formulas J2:J5 = E+G+I (0.25 points)
    try:
        col_j_pass = 0
        for row in range(2, 6):
            coord = f'J{row}'
            expected_f = EXPECTED_FORMULAS['J'].format(r=row)
            expected_v = EXPECTED_VALUES[coord]
            if check_cell_formula_or_value(ws, ws_data, coord, expected_f, expected_v):
                col_j_pass += 1
                print(f"PASS: {coord} — formula/value correct (expected {expected_v})")
            else:
                actual = ws[coord].value
                print(f"FAIL: {coord} — expected formula '{expected_f}' or value {expected_v}, found: {actual}")
        comp4_score = 0.25 * (col_j_pass / 4)
        if col_j_pass > 0:
            total_score += comp4_score
            print(f"PASS: Component 4 — Total Stockout Cost: {col_j_pass}/4 cells ({comp4_score:.2f} pts)")
        else:
            print(f"FAIL: Component 4 — Total Stockout Cost: 0/4 cells correct")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
