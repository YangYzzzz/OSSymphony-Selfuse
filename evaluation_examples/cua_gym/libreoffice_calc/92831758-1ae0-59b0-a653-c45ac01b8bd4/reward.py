"""
Reward Script: Pivot table with SUM of Quantity, SUM of TotalValue, and AvgUnitCost
Task ID: calc_pivot_095
Domain: libreoffice_calc
Scoring:
  Component 1 (0.15): PivotTable sheet exists (not present in initial)
  Component 2 (0.25): SUM of Quantity section — correct warehouse x category values
  Component 3 (0.25): SUM of TotalValue section — correct warehouse x category values
  Component 4 (0.20): AvgUnitCost section — correct calculated values (TotalValue/Quantity)
  Component 5 (0.15): Key ground truth values match (WH-Alpha/Raw Materials Qty=850, Value=42500, AvgUnitCost=50; Grand Total Qty=12000, Value=680000)
"""

import openpyxl
import os

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_095'

# Expected data from golden state exploration
# Warehouse order: WH-Alpha, WH-Beta, WH-Gamma, WH-Delta
# Category order: Raw Materials, Components, Finished Goods, Packaging

EXPECTED_QTY = {
    'WH-Alpha': {'Raw Materials': 850, 'Components': 770, 'Finished Goods': 730, 'Packaging': 600, 'Grand Total': 2950},
    'WH-Beta': {'Raw Materials': 970, 'Components': 830, 'Finished Goods': 690, 'Packaging': 660, 'Grand Total': 3150},
    'WH-Gamma': {'Raw Materials': 810, 'Components': 880, 'Finished Goods': 760, 'Packaging': 570, 'Grand Total': 3020},
    'WH-Delta': {'Raw Materials': 740, 'Components': 920, 'Finished Goods': 700, 'Packaging': 520, 'Grand Total': 2880},
    'Grand Total': {'Raw Materials': 3370, 'Components': 3400, 'Finished Goods': 2880, 'Packaging': 2350, 'Grand Total': 12000},
}

EXPECTED_VALUE = {
    'WH-Alpha': {'Raw Materials': 42500, 'Components': 50550, 'Finished Goods': 47750, 'Packaging': 27650, 'Grand Total': 168450},
    'WH-Beta': {'Raw Materials': 46150, 'Components': 54750, 'Finished Goods': 44950, 'Packaging': 30650, 'Grand Total': 176500},
    'WH-Gamma': {'Raw Materials': 38150, 'Components': 58250, 'Finished Goods': 49850, 'Packaging': 26150, 'Grand Total': 172400},
    'WH-Delta': {'Raw Materials': 34650, 'Components': 61050, 'Finished Goods': 45650, 'Packaging': 21300, 'Grand Total': 162650},
    'Grand Total': {'Raw Materials': 161450, 'Components': 224600, 'Finished Goods': 188200, 'Packaging': 105750, 'Grand Total': 680000},
}

EXPECTED_AVG = {
    'WH-Alpha': {'Raw Materials': 50.0, 'Components': 65.65, 'Finished Goods': 65.41, 'Packaging': 46.08, 'Grand Total': 57.1},
    'WH-Beta': {'Raw Materials': 47.58, 'Components': 65.96, 'Finished Goods': 65.14, 'Packaging': 46.44, 'Grand Total': 56.03},
    'WH-Gamma': {'Raw Materials': 47.10, 'Components': 66.19, 'Finished Goods': 65.59, 'Packaging': 45.88, 'Grand Total': 57.09},
    'WH-Delta': {'Raw Materials': 46.82, 'Components': 66.36, 'Finished Goods': 65.21, 'Packaging': 40.96, 'Grand Total': 56.48},
    'Grand Total': {'Raw Materials': 47.91, 'Components': 66.06, 'Finished Goods': 65.35, 'Packaging': 45.0, 'Grand Total': 56.67},
}

WAREHOUSES = ['WH-Alpha', 'WH-Beta', 'WH-Gamma', 'WH-Delta', 'Grand Total']
CATEGORIES = ['Raw Materials', 'Components', 'Finished Goods', 'Packaging', 'Grand Total']


def find_section_start(ws, label_substring):
    """Find the row where a section header containing label_substring starts."""
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val and label_substring.lower() in str(val).lower():
            return r
    return None


def parse_pivot_section(ws, header_row):
    """
    Parse a pivot section starting at header_row.
    header_row has column headers (Warehouse, cat1, cat2, ..., Grand Total).
    Data rows follow immediately after.
    Returns dict: {warehouse_name: {category: value, ...}, ...}
    """
    # Read column headers
    cols = {}
    for c in range(2, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        if h:
            cols[c] = str(h).strip()

    # Read data rows until empty
    data = {}
    for r in range(header_row + 1, ws.max_row + 1):
        wh = ws.cell(r, 1).value
        if wh is None:
            break
        wh_str = str(wh).strip()
        row_data = {}
        for c, cat in cols.items():
            val = ws.cell(r, c).value
            row_data[cat] = val
        data[wh_str] = row_data

    return data


def compare_section(actual, expected, tolerance=1.0, label=""):
    """
    Compare actual parsed section against expected.
    Returns (matches, total) count.
    """
    matches = 0
    total = 0
    for wh in expected:
        for cat in expected[wh]:
            total += 1
            exp_val = expected[wh][cat]
            act_val = actual.get(wh, {}).get(cat)
            if act_val is not None:
                try:
                    if abs(float(act_val) - float(exp_val)) <= tolerance:
                        matches += 1
                    else:
                        print(f"  MISMATCH {label} [{wh}][{cat}]: expected={exp_val}, actual={act_val}")
                except (ValueError, TypeError):
                    print(f"  TYPE_ERR {label} [{wh}][{cat}]: expected={exp_val}, actual={act_val}")
            else:
                print(f"  MISSING {label} [{wh}][{cat}]: expected={exp_val}")
    return matches, total


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: PivotTable sheet exists (0.15 points)
    # This is the key task-introduced change — initial has only 'InventoryFull'
    try:
        pivot_sheet_found = False
        for sn in wb.sheetnames:
            if 'pivot' in sn.lower():
                pivot_sheet_found = True
                pivot_sheet_name = sn
                break
        if pivot_sheet_found:
            print(f"PASS: Component 1 — PivotTable sheet found: '{pivot_sheet_name}' (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — No sheet with 'pivot' in name. Sheets: {wb.sheetnames}")
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print(f"REWARD: {total_score}")
        return total_score

    ws = wb[pivot_sheet_name]

    # Component 2: SUM of Quantity section correct (0.25 points)
    try:
        qty_header = find_section_start(ws, 'quantity')
        if qty_header is None:
            qty_header = find_section_start(ws, 'qty')
        if qty_header is not None:
            # The data header row is the next row after the section label
            data_header = qty_header + 1
            # Verify header row has 'Warehouse' label
            if ws.cell(data_header, 1).value and 'warehouse' in str(ws.cell(data_header, 1).value).lower():
                actual_qty = parse_pivot_section(ws, data_header)
            else:
                # Maybe the section label row IS the header row
                actual_qty = parse_pivot_section(ws, qty_header)

            matches, total_cells = compare_section(actual_qty, EXPECTED_QTY, tolerance=5.0, label="Qty")
            ratio = matches / total_cells if total_cells > 0 else 0
            pts = round(0.25 * ratio, 4)
            if ratio >= 0.8:
                print(f"PASS: Component 2 — SUM of Quantity: {matches}/{total_cells} cells match ({pts} pts)")
                total_score += pts
            else:
                print(f"PARTIAL: Component 2 — SUM of Quantity: {matches}/{total_cells} cells match ({pts} pts)")
                total_score += pts
        else:
            print(f"FAIL: Component 2 — No 'Quantity' section header found in PivotTable sheet")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: SUM of TotalValue section correct (0.25 points)
    try:
        val_header = find_section_start(ws, 'totalvalue')
        if val_header is None:
            val_header = find_section_start(ws, 'value')
        if val_header is not None:
            data_header = val_header + 1
            if ws.cell(data_header, 1).value and 'warehouse' in str(ws.cell(data_header, 1).value).lower():
                actual_val = parse_pivot_section(ws, data_header)
            else:
                actual_val = parse_pivot_section(ws, val_header)

            matches, total_cells = compare_section(actual_val, EXPECTED_VALUE, tolerance=50.0, label="Value")
            ratio = matches / total_cells if total_cells > 0 else 0
            pts = round(0.25 * ratio, 4)
            if ratio >= 0.8:
                print(f"PASS: Component 3 — SUM of TotalValue: {matches}/{total_cells} cells match ({pts} pts)")
                total_score += pts
            else:
                print(f"PARTIAL: Component 3 — SUM of TotalValue: {matches}/{total_cells} cells match ({pts} pts)")
                total_score += pts
        else:
            print(f"FAIL: Component 3 — No 'TotalValue' or 'Value' section header found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: AvgUnitCost section correct (0.20 points)
    try:
        avg_header = find_section_start(ws, 'avgunitcost')
        if avg_header is None:
            avg_header = find_section_start(ws, 'avg')
        if avg_header is None:
            avg_header = find_section_start(ws, 'unit cost')
        if avg_header is not None:
            data_header = avg_header + 1
            if ws.cell(data_header, 1).value and 'warehouse' in str(ws.cell(data_header, 1).value).lower():
                actual_avg = parse_pivot_section(ws, data_header)
            else:
                actual_avg = parse_pivot_section(ws, avg_header)

            matches, total_cells = compare_section(actual_avg, EXPECTED_AVG, tolerance=1.0, label="AvgCost")
            ratio = matches / total_cells if total_cells > 0 else 0
            pts = round(0.20 * ratio, 4)
            if ratio >= 0.8:
                print(f"PASS: Component 4 — AvgUnitCost: {matches}/{total_cells} cells match ({pts} pts)")
                total_score += pts
            else:
                print(f"PARTIAL: Component 4 — AvgUnitCost: {matches}/{total_cells} cells match ({pts} pts)")
                total_score += pts
        else:
            print(f"FAIL: Component 4 — No 'AvgUnitCost' section header found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Key ground truth values (0.15 points)
    # Verify specific values mentioned in task context:
    # WH-Alpha/Raw Materials: Qty=850, Value=42500, AvgUnitCost=50
    # Grand Total Qty=12000, Value=680000
    try:
        gt_checks_passed = 0
        gt_total = 5

        # Check WH-Alpha / Raw Materials Qty = 850
        qty_header_row = find_section_start(ws, 'quantity')
        if qty_header_row:
            dh = qty_header_row + 1
            if ws.cell(dh, 1).value and 'warehouse' in str(ws.cell(dh, 1).value).lower():
                actual_qty_sec = parse_pivot_section(ws, dh)
            else:
                actual_qty_sec = parse_pivot_section(ws, qty_header_row)
            alpha_rm_qty = actual_qty_sec.get('WH-Alpha', {}).get('Raw Materials')
            if alpha_rm_qty is not None and abs(float(alpha_rm_qty) - 850) <= 1:
                gt_checks_passed += 1
                print(f"  GT-CHECK: WH-Alpha/RawMaterials Qty=850 -> OK ({alpha_rm_qty})")
            else:
                print(f"  GT-CHECK: WH-Alpha/RawMaterials Qty=850 -> FAIL ({alpha_rm_qty})")

            grand_qty = actual_qty_sec.get('Grand Total', {}).get('Grand Total')
            if grand_qty is not None and abs(float(grand_qty) - 12000) <= 1:
                gt_checks_passed += 1
                print(f"  GT-CHECK: Grand Total Qty=12000 -> OK ({grand_qty})")
            else:
                print(f"  GT-CHECK: Grand Total Qty=12000 -> FAIL ({grand_qty})")

        # Check WH-Alpha / Raw Materials Value = 42500
        val_header_row = find_section_start(ws, 'totalvalue')
        if val_header_row is None:
            val_header_row = find_section_start(ws, 'value')
        if val_header_row:
            dh = val_header_row + 1
            if ws.cell(dh, 1).value and 'warehouse' in str(ws.cell(dh, 1).value).lower():
                actual_val_sec = parse_pivot_section(ws, dh)
            else:
                actual_val_sec = parse_pivot_section(ws, val_header_row)
            alpha_rm_val = actual_val_sec.get('WH-Alpha', {}).get('Raw Materials')
            if alpha_rm_val is not None and abs(float(alpha_rm_val) - 42500) <= 10:
                gt_checks_passed += 1
                print(f"  GT-CHECK: WH-Alpha/RawMaterials Value=42500 -> OK ({alpha_rm_val})")
            else:
                print(f"  GT-CHECK: WH-Alpha/RawMaterials Value=42500 -> FAIL ({alpha_rm_val})")

            grand_val = actual_val_sec.get('Grand Total', {}).get('Grand Total')
            if grand_val is not None and abs(float(grand_val) - 680000) <= 10:
                gt_checks_passed += 1
                print(f"  GT-CHECK: Grand Total Value=680000 -> OK ({grand_val})")
            else:
                print(f"  GT-CHECK: Grand Total Value=680000 -> FAIL ({grand_val})")

        # Check WH-Alpha / Raw Materials AvgUnitCost = 50
        avg_header_row = find_section_start(ws, 'avgunitcost')
        if avg_header_row is None:
            avg_header_row = find_section_start(ws, 'avg')
        if avg_header_row is None:
            avg_header_row = find_section_start(ws, 'unit cost')
        if avg_header_row:
            dh = avg_header_row + 1
            if ws.cell(dh, 1).value and 'warehouse' in str(ws.cell(dh, 1).value).lower():
                actual_avg_sec = parse_pivot_section(ws, dh)
            else:
                actual_avg_sec = parse_pivot_section(ws, avg_header_row)
            alpha_rm_avg = actual_avg_sec.get('WH-Alpha', {}).get('Raw Materials')
            if alpha_rm_avg is not None and abs(float(alpha_rm_avg) - 50.0) <= 1.0:
                gt_checks_passed += 1
                print(f"  GT-CHECK: WH-Alpha/RawMaterials AvgUnitCost=50 -> OK ({alpha_rm_avg})")
            else:
                print(f"  GT-CHECK: WH-Alpha/RawMaterials AvgUnitCost=50 -> FAIL ({alpha_rm_avg})")

        pts = round(0.15 * (gt_checks_passed / gt_total), 4)
        if gt_checks_passed >= 4:
            print(f"PASS: Component 5 — Ground truth values: {gt_checks_passed}/{gt_total} match ({pts} pts)")
        else:
            print(f"PARTIAL: Component 5 — Ground truth values: {gt_checks_passed}/{gt_total} match ({pts} pts)")
        total_score += pts
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
