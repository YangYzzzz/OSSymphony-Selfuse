"""
Initial Setup: Inventory data across multiple warehouses
Task ID: calc_pivot_095
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_095'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


# ---- Data Design ----
# We need 300 rows distributed across 4 warehouses x 4 categories = 16 groups.
# Ground truth constraints:
#   WH-Alpha / Raw Materials: Qty=850, Value=42500 => AvgUnitCost=50
#   Grand total: Qty=12000, Value=680000
#
# We'll design explicit group totals that satisfy these constraints.

WAREHOUSES = ['WH-Alpha', 'WH-Beta', 'WH-Gamma', 'WH-Delta']
CATEGORIES = ['Raw Materials', 'Components', 'Finished Goods', 'Packaging']

# (warehouse_idx, category_idx): (total_qty, total_value, num_rows)
# Design so that grand total Qty=12000, Value=680000
# WH-Alpha/Raw Materials: Qty=850, Value=42500
GROUP_SPEC = {
    # WH-Alpha
    (0, 0): (850,  42500,  18),   # Raw Materials
    (0, 1): (770,  50550,  18),   # Components
    (0, 2): (730,  47750,  18),   # Finished Goods
    (0, 3): (600,  27650,  18),   # Packaging
    # WH-Beta
    (1, 0): (970,  46150,  19),   # Raw Materials
    (1, 1): (830,  54750,  19),   # Components
    (1, 2): (690,  44950,  19),   # Finished Goods
    (1, 3): (660,  30650,  19),   # Packaging
    # WH-Gamma
    (2, 0): (810,  38150,  19),   # Raw Materials
    (2, 1): (880,  58250,  19),   # Components
    (2, 2): (760,  49850,  19),   # Finished Goods
    (2, 3): (570,  26150,  18),   # Packaging
    # WH-Delta
    (3, 0): (740,  34650,  19),   # Raw Materials
    (3, 1): (920,  61050,  19),   # Components
    (3, 2): (700,  45650,  19),   # Finished Goods
    (3, 3): (520,  21300,  20),   # Packaging
}

# Verify totals
total_qty = sum(v[0] for v in GROUP_SPEC.values())
total_val = sum(v[1] for v in GROUP_SPEC.values())
total_rows = sum(v[2] for v in GROUP_SPEC.values())
assert total_qty == 12000, f"Qty total mismatch: {total_qty}"
assert total_val == 680000, f"Value total mismatch: {total_val}"
assert total_rows == 300, f"Row total mismatch: {total_rows}"

# Product name pools by category
PRODUCT_NAMES = {
    'Raw Materials': [
        'Steel Sheet Grade A', 'Copper Wire 2mm', 'Aluminum Bar 6061',
        'PVC Granules', 'Carbon Fiber Roll', 'Stainless Rod 304',
        'Brass Tube 10mm', 'Rubber Sheet Natural', 'Silicone Compound',
        'Glass Fiber Mat', 'Titanium Strip', 'Nylon Pellets',
        'Zinc Alloy Ingot', 'Polyethylene Resin', 'Ceramic Powder',
        'Epoxy Resin Base', 'Graphite Block', 'Acrylic Sheet 3mm',
        'Polycarbonate Pellets', 'Teflon Rod',
    ],
    'Components': [
        'Bearing SKF-6205', 'Circuit Board PCB-A', 'Capacitor 100uF',
        'Motor Brushless DC', 'Sensor Temp PT100', 'Relay 24V DPDT',
        'Connector USB-C', 'LED Module RGB', 'Transformer 12V',
        'Switch Toggle SPST', 'Resistor Pack 1K', 'Diode Bridge Rect',
        'Fan Cooling 80mm', 'Fuse Holder 5A', 'Battery Cell 3.7V',
        'Cable Harness Kit', 'Solenoid Valve 12V', 'Gear Set Nylon',
        'Shaft Coupling Flex', 'Encoder Rotary 360',
    ],
    'Finished Goods': [
        'Smart Thermostat Pro', 'LED Panel Light 60W', 'Air Purifier Mini',
        'Power Bank 20000mAh', 'Wireless Charger Pad', 'Bluetooth Speaker V3',
        'USB Hub 7-Port', 'Desk Lamp Adjustable', 'Digital Scale 5kg',
        'Electric Kettle 1.5L', 'Robot Vacuum Basic', 'Smart Plug WiFi',
        'Portable Fan USB', 'Night Light Sensor', 'Cable Organizer Box',
        'Monitor Stand Riser', 'Keyboard Mechanical', 'Mouse Wireless Ergo',
        'Webcam HD 1080p', 'Headset Noise Cancel',
    ],
    'Packaging': [
        'Cardboard Box 30x20', 'Bubble Wrap Roll 50m', 'Shrink Film 100m',
        'Packing Tape Clear', 'Foam Insert Custom', 'Poly Bag 12x15',
        'Corrugated Divider', 'Label Roll Thermal', 'Strapping Band PP',
        'Pallet Wrap 500mm', 'Edge Protector Card', 'Tissue Paper White',
        'Zip Lock Bag 8x10', 'Mailing Envelope Pad', 'Box Filler Peanuts',
        'Desiccant Silica Gel', 'Seal Sticker Round', 'Gift Box Premium',
        'Cling Film Food Grade', 'Kraft Paper Roll',
    ],
}


def distribute_values(total, num_items, min_val=5):
    """Distribute total into num_items random positive integers summing to total."""
    if num_items == 1:
        return [total]
    # Start with minimum values
    values = [min_val] * num_items
    remaining = total - min_val * num_items
    if remaining < 0:
        raise ValueError(f"Cannot distribute {total} into {num_items} items with min {min_val}")
    # Distribute remaining randomly
    for i in range(remaining):
        idx = random.randint(0, num_items - 1)
        values[idx] += 1
    random.shuffle(values)
    return values


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'InventoryFull'

    # Headers
    headers = ['SKU', 'ProductName', 'Warehouse', 'Category', 'Quantity', 'TotalValue']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14

    # Generate data rows
    rows = []
    sku_counter = 1

    for (wi, ci), (grp_qty, grp_val, num_rows) in GROUP_SPEC.items():
        warehouse = WAREHOUSES[wi]
        category = CATEGORIES[ci]
        product_pool = PRODUCT_NAMES[category]

        qtys = distribute_values(grp_qty, num_rows, min_val=10)
        vals = distribute_values(grp_val, num_rows, min_val=50)

        for i in range(num_rows):
            product = product_pool[i % len(product_pool)]
            rows.append([sku_counter, product, warehouse, category, qtys[i], vals[i]])
            sku_counter += 1

    # Shuffle rows so data isn't grouped by warehouse/category
    random.shuffle(rows)

    # Re-assign SKUs sequentially after shuffle
    for idx, row in enumerate(rows):
        row[0] = idx + 1

    # Write data
    for r, row_data in enumerate(rows, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c in (5, 6):  # Quantity and TotalValue
                cell.number_format = '#,##0'

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = f'A1:F{len(rows) + 1}'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
