"""
Initial Setup: Employee satisfaction survey data for stacked bar chart creation
Task ID: calc_gcp_076
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_076'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'SurveyViz'

    # Headers
    headers = ['Question', 'StronglyDisagree', 'Disagree', 'Neutral', 'Agree', 'StronglyAgree']
    header_font = Font(name='Arial', size=11, bold=True)
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_font_white = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Survey data - percentages that sum to 100% per row
    # Categories: Work-Life Balance, Career Growth, Compensation, Management, Team Culture
    data = [
        ['Work-Life Balance',   8,  12, 15, 35, 30],
        ['Career Growth',      12,  18, 22, 28, 20],
        ['Compensation',       15,  16, 20, 30, 19],
        ['Management',         10,  14, 18, 33, 25],
        ['Team Culture',        5,   8, 12, 32, 43],
    ]

    data_font = Font(name='Arial', size=11)
    pct_format = '0"%"'

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = thin_border
            if c == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.number_format = pct_format
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Column widths
    ws.column_dimensions['A'].width = 22
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col_letter].width = 18

    # Row height for header
    ws.row_dimensions[1].height = 30

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
