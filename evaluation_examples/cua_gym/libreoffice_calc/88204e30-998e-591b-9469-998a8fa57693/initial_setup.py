"""
Initial Setup: Dual chart creation task - Inbound and Outbound shipment tables
Task ID: osworld_calc_dual_chart_separate_tables_005
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_dual_chart_separate_tables_005'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Shipments'

    # --- Inbound Shipments Table ---
    # Row 1: Column headers for inbound table (Route in A, Volume in B)
    ws.cell(row=1, column=1, value='Route')
    ws.cell(row=1, column=2, value='Volume')
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=2).font = Font(bold=True)

    # Rows 2-10: Inbound data (9 rows of data)
    inbound_data = [
        ('New York - Chicago', 4820),
        ('Los Angeles - Dallas', 3610),
        ('Chicago - Miami', 5270),
        ('Houston - Atlanta', 2940),
        ('Phoenix - Denver', 1870),
        ('Seattle - Portland', 3350),
        ('Boston - Philadelphia', 4120),
        ('San Francisco - Las Vegas', 2760),
        ('Detroit - Cleveland', 3490),
    ]
    for row_idx, (route, volume) in enumerate(inbound_data, start=2):
        ws.cell(row=row_idx, column=1, value=route)
        ws.cell(row=row_idx, column=2, value=volume)

    # Row 11: blank separator row

    # --- Outbound Shipments Table ---
    # Row 12: Column headers for outbound table (Route in D, Volume in E)
    ws.cell(row=12, column=4, value='Route')
    ws.cell(row=12, column=5, value='Volume')
    ws.cell(row=12, column=4).font = Font(bold=True)
    ws.cell(row=12, column=5).font = Font(bold=True)

    # Rows 13-21: Outbound data (9 rows of data)
    outbound_data = [
        ('Chicago - New York', 3980),
        ('Dallas - Los Angeles', 4430),
        ('Miami - Chicago', 2870),
        ('Atlanta - Houston', 3650),
        ('Denver - Phoenix', 1990),
        ('Portland - Seattle', 2810),
        ('Philadelphia - Boston', 3720),
        ('Las Vegas - San Francisco', 4160),
        ('Cleveland - Detroit', 2930),
    ]
    for row_idx, (route, volume) in enumerate(outbound_data, start=13):
        ws.cell(row=row_idx, column=4, value=route)
        ws.cell(row=row_idx, column=5, value=volume)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 12

    # NOTE: No charts are added to the initial file — the task is to create them

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
