"""
Reward Script: Calculate time-to-hire metrics for recruiting dashboard
Task ID: calc_hr_time_to_hire_036
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.40): NETWORKDAYS/IF formulas in F2:F73 for all rows
  - Component 2 (0.35): Summary section I1:J7 with dept names (bold headers) and AVERAGEIF formulas
  - Component 3 (0.25): Conditional formatting on A2:G73 highlighting rows with >45 working days in orange
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_time_to_hire_036'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify 'Recruiting' sheet exists (precondition gate)
    if 'Recruiting' not in wb.sheetnames:
        print("CRITICAL: 'Recruiting' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Recruiting']

    # -----------------------------------------------------------------------
    # Component 1: NETWORKDAYS/IF formulas in F2:F73 (0.40 points)
    # Each row should have: =IF(G{row}="Filled",NETWORKDAYS(D{row},E{row},"")
    # This FAILS on initial (column F is empty) and PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        correct_formula_count = 0
        total_formula_rows = 72  # rows 2-73

        for row in range(2, 74):
            val = ws.cell(row=row, column=6).value  # column F
            if val is None:
                continue
            if not isinstance(val, str):
                continue
            # Check that the formula uses IF with "Filled" condition and NETWORKDAYS
            val_upper = val.upper().replace(' ', '')
            # Expected pattern: =IF(G{row}="Filled",NETWORKDAYS(D{row},E{row},"")
            # Normalize and check key parts
            has_if = 'IF(' in val_upper
            has_filled = '"FILLED"' in val_upper
            has_networkdays = 'NETWORKDAYS(' in val_upper
            # Check that the correct row's D and E columns are referenced
            d_ref = f'D{row}'
            e_ref = f'E{row}'
            g_ref = f'G{row}'
            has_correct_refs = (
                d_ref.upper() in val_upper and
                e_ref.upper() in val_upper and
                g_ref.upper() in val_upper
            )
            if has_if and has_filled and has_networkdays and has_correct_refs:
                correct_formula_count += 1

        ratio = correct_formula_count / total_formula_rows
        if ratio >= 0.95:
            print(f"PASS: Component 1 — NETWORKDAYS/IF formulas in F2:F73 ({correct_formula_count}/{total_formula_rows} correct) (0.4 pts)")
            total_score += 0.40
        elif ratio >= 0.5:
            partial = round(0.40 * ratio, 2)
            print(f"PARTIAL: Component 1 — {correct_formula_count}/{total_formula_rows} NETWORKDAYS/IF formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {correct_formula_count}/{total_formula_rows} rows have correct NETWORKDAYS/IF formulas")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Summary section in I1:J7 with headers and AVERAGEIF formulas
    # I1='Department' (bold), J1='Avg Days to Hire' (bold),
    # I2:I7 = department names, J2:J7 = AVERAGEIF formulas
    # This FAILS on initial (I/J columns are empty) and PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        summary_score = 0.0

        # Check I1 header
        i1_val = ws['I1'].value
        i1_bold = ws['I1'].font.bold
        if i1_val == 'Department' and i1_bold:
            summary_score += 0.05
            print(f"PASS: I1='Department' (bold)")
        elif i1_val == 'Department':
            summary_score += 0.025
            print(f"PARTIAL: I1='Department' but not bold")
        else:
            print(f"FAIL: I1 expected 'Department' (bold), got {repr(i1_val)}, bold={i1_bold}")

        # Check J1 header
        j1_val = ws['J1'].value
        j1_bold = ws['J1'].font.bold
        if j1_val == 'Avg Days to Hire' and j1_bold:
            summary_score += 0.05
            print(f"PASS: J1='Avg Days to Hire' (bold)")
        elif j1_val == 'Avg Days to Hire':
            summary_score += 0.025
            print(f"PARTIAL: J1='Avg Days to Hire' but not bold")
        else:
            print(f"FAIL: J1 expected 'Avg Days to Hire' (bold), got {repr(j1_val)}, bold={j1_bold}")

        # Check department names in I2:I7
        expected_depts = ['Engineering', 'Marketing', 'Finance', 'Operations', 'HR', 'Sales']
        depts_found = []
        for i, row in enumerate(range(2, 8)):
            i_val = ws.cell(row=row, column=9).value
            depts_found.append(i_val)

        # Check all 6 departments are present (order matters per context)
        correct_depts = sum(1 for i, d in enumerate(depts_found) if d == expected_depts[i])
        if correct_depts == 6:
            summary_score += 0.10
            print(f"PASS: I2:I7 all 6 department names correct: {depts_found}")
        elif correct_depts >= 4:
            summary_score += 0.05
            print(f"PARTIAL: {correct_depts}/6 department names correct in I2:I7")
        else:
            print(f"FAIL: Only {correct_depts}/6 department names correct. Found: {depts_found}")

        # Check AVERAGEIF formulas in J2:J7
        correct_avgs = 0
        for row in range(2, 8):
            j_val = ws.cell(row=row, column=10).value
            if j_val is None:
                continue
            if not isinstance(j_val, str):
                continue
            j_upper = j_val.upper().replace(' ', '')
            # Expected: =AVERAGEIF($B$2:$B$73,I{row},$F$2:$F$73)
            has_averageif = 'AVERAGEIF(' in j_upper
            has_b_range = '$B$2:$B$73' in j_upper or 'B2:B73' in j_upper
            has_f_range = '$F$2:$F$73' in j_upper or 'F2:F73' in j_upper
            has_i_ref = f'I{row}' in j_val.upper()
            if has_averageif and has_b_range and has_f_range and has_i_ref:
                correct_avgs += 1

        if correct_avgs == 6:
            summary_score += 0.15
            print(f"PASS: J2:J7 all 6 AVERAGEIF formulas correct")
        elif correct_avgs >= 3:
            partial = round(0.15 * correct_avgs / 6, 3)
            summary_score += partial
            print(f"PARTIAL: {correct_avgs}/6 AVERAGEIF formulas correct in J2:J7 ({partial} pts)")
        else:
            print(f"FAIL: Only {correct_avgs}/6 AVERAGEIF formulas correct in J2:J7")

        if summary_score >= 0.30:
            print(f"PASS: Component 2 — Summary section complete ({summary_score:.3f}/0.35 pts awarded)")
        else:
            print(f"PARTIAL/FAIL: Component 2 — Summary section incomplete ({summary_score:.3f}/0.35 pts)")
        total_score += min(summary_score, 0.35)

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Conditional formatting on A2:G73 highlighting rows with
    # >45 working days in orange (FFFF6600)
    # Formula: AND($F2>45,$G2="Filled") → fill fgColor = FFFF6600
    # This FAILS on initial (no CF rules present) and PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        found_cf = False
        found_correct_color = False
        found_correct_formula = False
        found_correct_range = False

        for cf_range_obj in ws.conditional_formatting:
            # Check if this CF covers A2:G73 (or a superset/subset that includes A2:G73)
            cf_str = str(cf_range_obj)
            range_ok = 'A2:G73' in cf_str or ('A2' in cf_str and 'G73' in cf_str)

            for rule in ws.conditional_formatting[cf_range_obj]:
                found_cf = True
                # Check formula contains F>45 and G="Filled"
                formula = getattr(rule, 'formula', None)
                if formula:
                    formula_str = ' '.join(formula).upper().replace(' ', '')
                    has_f_gt_45 = ('$F' in formula_str or 'F2' in formula_str.upper()) and '>45' in formula_str
                    has_g_filled = '"FILLED"' in formula_str
                    if has_f_gt_45 and has_g_filled:
                        found_correct_formula = True

                # Check fill color is orange FFFF6600
                if hasattr(rule, 'dxf') and rule.dxf is not None:
                    dxf = rule.dxf
                    if hasattr(dxf, 'fill') and dxf.fill is not None:
                        try:
                            color_rgb = dxf.fill.fgColor.rgb
                            # Accept FFFF6600 (orange) or FF6600 with FF prefix
                            if color_rgb in ('FFFF6600', 'FF6600'):
                                found_correct_color = True
                        except Exception:
                            pass

                if range_ok:
                    found_correct_range = True

        if found_correct_formula and found_correct_color and found_cf:
            print(f"PASS: Component 3 — Conditional formatting with orange (FFFF6600) for rows where F>45 AND G='Filled' (0.25 pts)")
            total_score += 0.25
        elif found_cf and found_correct_formula:
            print(f"PARTIAL: Component 3 — CF formula correct but color may not be FFFF6600 (0.1 pts)")
            total_score += 0.10
        elif found_cf and found_correct_color:
            print(f"PARTIAL: Component 3 — CF color correct but formula may not match (0.1 pts)")
            total_score += 0.10
        elif found_cf:
            print(f"PARTIAL: Component 3 — CF exists but formula/color incorrect (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 3 — No conditional formatting found on worksheet")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.3f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
