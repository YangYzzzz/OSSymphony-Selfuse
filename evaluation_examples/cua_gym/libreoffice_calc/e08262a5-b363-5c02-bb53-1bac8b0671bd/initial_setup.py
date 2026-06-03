"""
Initial Setup: HR Recruiting Dashboard - Time to Hire Metrics
Task ID: calc_hr_time_to_hire_036
Domain: libreoffice_calc
Creates: calc_hr_time_to_hire_036_initial.xlsx
"""

import os
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_time_to_hire_036'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Recruiting ---
    ws = wb.active
    ws.title = 'Recruiting'

    # Headers
    headers = ['Req ID', 'Department', 'Job Title', 'Posted Date',
               'Offer Accepted Date', 'Working Days to Hire', 'Status']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Departments
    departments = ['Engineering', 'Marketing', 'Finance', 'Operations', 'HR', 'Sales']

    # Job titles per department
    job_titles = {
        'Engineering': ['Software Engineer', 'Senior Developer', 'DevOps Engineer',
                        'QA Engineer', 'Data Engineer', 'Backend Developer',
                        'Frontend Developer', 'ML Engineer', 'Platform Engineer',
                        'Security Engineer', 'Cloud Architect', 'Systems Analyst'],
        'Marketing':  ['Marketing Analyst', 'Brand Manager', 'Content Strategist',
                       'SEO Specialist', 'Campaign Manager', 'Digital Marketer',
                       'Product Marketing Manager', 'Social Media Manager',
                       'Growth Hacker', 'UX Researcher', 'Market Research Analyst'],
        'Finance':    ['Financial Analyst', 'Senior Accountant', 'Budget Analyst',
                       'Treasury Analyst', 'Payroll Specialist', 'Tax Associate',
                       'Finance Manager', 'Controller', 'Audit Associate',
                       'Credit Analyst', 'Investment Analyst'],
        'Operations': ['Operations Analyst', 'Supply Chain Manager', 'Logistics Coordinator',
                       'Process Improvement Specialist', 'Procurement Specialist',
                       'Operations Manager', 'Facilities Coordinator', 'Inventory Analyst',
                       'Quality Assurance Specialist', 'Business Analyst'],
        'HR':         ['HR Generalist', 'Recruiter', 'HR Business Partner',
                       'Compensation Analyst', 'Training Specialist', 'HR Manager',
                       'Benefits Coordinator', 'Talent Acquisition Specialist',
                       'HRIS Analyst', 'Employee Relations Specialist'],
        'Sales':      ['Account Executive', 'Sales Development Rep', 'Senior AE',
                       'Sales Manager', 'Customer Success Manager', 'Inside Sales Rep',
                       'Regional Sales Director', 'Business Development Rep',
                       'Key Account Manager', 'Territory Manager', 'Sales Analyst'],
    }

    # Candidate names pool
    first_names = ['Sarah', 'Marcus', 'Jennifer', 'David', 'Emily', 'James', 'Ashley',
                   'Robert', 'Michelle', 'Kevin', 'Laura', 'Brian', 'Rachel', 'Daniel',
                   'Megan', 'Tyler', 'Amanda', 'Christopher', 'Jessica', 'Andrew',
                   'Natalie', 'Steven', 'Samantha', 'Matthew', 'Lauren', 'Joshua',
                   'Kayla', 'Justin', 'Brittany', 'Nathan', 'Amber', 'Eric', 'Heather',
                   'Patrick', 'Christina', 'Ryan', 'Stephanie', 'Brandon', 'Rebecca',
                   'Adam', 'Nicole', 'Aaron', 'Elizabeth', 'Jonathan', 'Danielle',
                   'Carlos', 'Priya', 'Wei', 'Amara', 'Tariq']
    last_names = ['Chen', 'Johnson', 'Williams', 'Thompson', 'Martinez', 'Davis',
                  'Wilson', 'Anderson', 'Garcia', 'Taylor', 'Brown', 'Lee', 'Harris',
                  'Jackson', 'White', 'Lewis', 'Robinson', 'Walker', 'Hall', 'Allen',
                  'Young', 'King', 'Wright', 'Scott', 'Green', 'Baker', 'Adams',
                  'Nelson', 'Carter', 'Mitchell', 'Patel', 'Kim', 'Singh', 'Nguyen',
                  'Lopez', 'Hill', 'Moore', 'Turner', 'Phillips', 'Campbell']

    # Status distribution: ~55% Filled, ~25% Open, ~20% Cancelled
    statuses = (
        ['Filled'] * 40 +
        ['Open'] * 18 +
        ['Cancelled'] * 14
    )

    # Base date for postings: spread over 2024-2025
    base_date = date(2024, 1, 8)

    import random
    random.seed(42)

    records = []
    name_idx = 0
    for i in range(72):
        dept = departments[i % len(departments)]
        titles = job_titles[dept]
        title = titles[i % len(titles)]

        req_id = f'REQ-{1001 + i:04d}'

        # Posted date: spread across 18 months
        posted = base_date + timedelta(days=random.randint(0, 540))

        # Status
        status = statuses[i % len(statuses)]

        # For Filled: offer date = posted + 20-80 calendar days
        # For Open/Cancelled: offer date is empty
        if status == 'Filled':
            calendar_days = random.randint(28, 95)
            offer_date = posted + timedelta(days=calendar_days)
        else:
            offer_date = None

        records.append((req_id, dept, title, posted, offer_date, None, status))
        name_idx += 1

    # Shuffle to mix departments
    random.shuffle(records)

    for row_idx, (req_id, dept, title, posted, offer_date, _, status) in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=req_id)
        ws.cell(row=row_idx, column=2, value=dept)
        ws.cell(row=row_idx, column=3, value=title)

        d_cell = ws.cell(row=row_idx, column=4, value=posted)
        d_cell.number_format = 'yyyy-mm-dd'

        if offer_date:
            e_cell = ws.cell(row=row_idx, column=5, value=offer_date)
            e_cell.number_format = 'yyyy-mm-dd'
        else:
            ws.cell(row=row_idx, column=5, value=None)

        # Column F (Working Days to Hire) is intentionally left EMPTY
        ws.cell(row=row_idx, column=6, value=None)

        ws.cell(row=row_idx, column=7, value=status)

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Recruiting')
    print(f'  Rows: 73 (1 header + 72 data rows)')
    print(f'  Column F (Working Days to Hire): EMPTY (agent must fill)')
    print(f'  No conditional formatting, no summary section')


create_initial()
