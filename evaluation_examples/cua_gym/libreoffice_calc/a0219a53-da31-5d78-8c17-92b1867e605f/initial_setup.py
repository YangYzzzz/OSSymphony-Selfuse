"""
Initial Setup: Sales workbook with customer data, commission rates, and summary
Task ID: calc_sales_customer_named_ranges_025
Domain: libreoffice_calc

Creates a workbook WITHOUT named ranges. The task is to define named ranges
(CustomerData, PlatinumList, CommissionRates) and update Summary formulas to use them.
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_customer_named_ranges_025'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # ----------------------------------------------------------------
    # Sheet 1: Customers (A2:F201 = 200 rows of customer data)
    # ----------------------------------------------------------------
    ws_cust = wb.active
    ws_cust.title = 'Customers'

    headers = ['CustomerID', 'Name', 'Region', 'Revenue', 'Tier', 'JoinDate']
    for col, h in enumerate(headers, 1):
        cell = ws_cust.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Realistic customer data: ~20 Platinum, ~45 Gold, rest Silver/Bronze
    platinum_names = [
        'Meridian Technologies', 'Apex Global Holdings', 'Crestview Capital',
        'Summit Ventures Inc', 'Pinnacle Industries', 'Horizon Dynamics',
        'Nexus Corporation', 'Vanguard Solutions', 'Sterling Enterprises',
        'Titan Financial Group', 'Catalyst Systems', 'Prism Analytics',
        'Vertex Networks', 'Eclipse Consulting', 'Aurora Partners',
        'Zenith Manufacturing', 'Polar Star Holdings', 'Crown Ridge Corp',
        'Pacific Rim Traders', 'Atlantic Bridge LLC'
    ]
    gold_names = [
        'Blue Ridge Exports', 'Cedar Valley Logistics', 'Diamond Path Group',
        'Eagle Eye Distribution', 'Forest Glen Retail', 'Golden Gate Imports',
        'Harbor View Services', 'Inland Empire Corp', 'Jade River Trading',
        'Keystone Solutions', 'Lakeside Partners', 'Maple Grove Industries',
        'Northern Lights Media', 'Oakwood Consulting', 'Pacific Crest Ventures',
        'Quartz Peak Holdings', 'Redwood Enterprises', 'Silver Creek Tech',
        'Thunder Valley Farms', 'Union Square Capital', 'Valley Forge Ltd',
        'Waverly Distributors', 'Xenon Digital Media', 'Yellow Stone Energy',
        'Zephyr Aerospace', 'Amber Cove Holdings', 'Birchwood Partners',
        'Cobalt Blue Systems', 'Driftwood Marketing', 'Ember Falls Group',
        'Fern Hollow Retail', 'Granite Peak Services', 'Heatherwood Corp',
        'Iron Bridge Trading', 'Juniper Hill Capital', 'Kingsway Solutions',
        'Lakeview Analytics', 'Mountainside Ventures', 'Nordic Crown LLC',
        'Orchard Hill Partners', 'Persimmon Grove Inc', 'Quail Run Holdings',
        'Riverstone Consulting', 'Sagebrush Industries', 'Tumbleweed Media'
    ]
    silver_names = [
        'Acorn Valley Farms', 'Brook Side Retail', 'Canyon Road Services',
        'Dune Ridge Corp', 'Eastern Shore Trading', 'Fairway Solutions',
        'Glenhurst Partners', 'Hillcrest Logistics', 'Ironwood Distributors',
        'Jasmine Court Holdings', 'Kilimanjaro Ventures', 'Lemon Tree Group',
        'Maplewood Services', 'Nighthawk Capital', 'Olive Branch Corp',
        'Pebble Creek Trading', 'Quarrystone Partners', 'Rustic Ridge Farms',
        'Sandstone Media', 'Timberland Solutions', 'Umbra Holdings',
        'Vineyard Hill Retail', 'Westbrook Industries', 'Xerxes Systems',
        'Yellowwood Corp', 'Zinnia Gardens Ltd', 'Alder Grove Partners',
        'Bayshore Consulting', 'Cliffside Trading', 'Dusty Trail Corp',
        'Evergreen Ventures', 'Foxglove Holdings', 'Glacier Ridge Group',
        'Huckleberry Farms', 'Indigo Bay Trading', 'Junglewood Corp',
        'Kettle Creek Solutions', 'Larkspur Holdings', 'Moonrise Partners',
        'Narwhal Capital', 'Osprey Ridge LLC', 'Primrose Path Corp',
        'Quartzite Holdings', 'Redstone Media', 'Silverbell Ventures',
        'Thornwood Retail', 'Underhill Partners', 'Vespera Holdings',
        'Windhaven Corp', 'Xenolith Systems', 'Yarrow Fields Inc',
        'Zebra Creek Trading', 'Arroyo Partners', 'Bayou Star Holdings',
        'Clearwater Corp', 'Driftnet Solutions', 'Elkhorn Ventures',
        'Frostbite Media', 'Granite Falls Group', 'Highbury Retail',
        'Ironclad Partners', 'Juniper Crest Corp', 'Knollwood Holdings',
        'Limestone Ridge LLC', 'Misty Vale Trading'
    ]
    bronze_names = [
        'Acacia Ridge Corp', 'Bayside Ventures', 'Cedarwood Trading',
        'Desert Rose Solutions', 'Eastern Gate Holdings', 'Fernwood Partners',
        'Graystone Corp', 'Hazel Crest Group', 'Island Peak Trading',
        'Jade Forest Holdings', 'Kendall Ridge Corp', 'Lava Rock Ventures',
        'Mosswood Solutions', 'Nightshade Corp', 'Oceanside Partners',
        'Pinecone Holdings', 'Quaking Aspen LLC', 'Riverwood Media',
        'Saltmarsh Trading', 'Tidewater Holdings', 'Upland Corp',
        'Verdant Hills Partners', 'Willow Creek Ventures', 'Xeric Solutions',
        'Yosemite Holdings', 'Zinfandel Corp', 'Ash Grove Partners',
        'Bluebell Trading', 'Copperhead Holdings', 'Dragonfly Corp'
    ]

    regions = ['North America', 'Europe', 'Asia Pacific', 'Latin America', 'Middle East']

    import random
    random.seed(42)
    import datetime

    rows = []
    cid = 1001

    # Add Platinum (~20 rows) - high revenue
    for name in platinum_names:
        revenue = random.randint(850000, 2500000)
        region = random.choice(regions)
        days_ago = random.randint(365*3, 365*8)
        join_date = datetime.date(2025, 3, 4) - datetime.timedelta(days=days_ago)
        rows.append([f'C{cid:04d}', name, region, revenue, 'Platinum', join_date.strftime('%Y-%m-%d')])
        cid += 1

    # Add Gold (~45 rows) - medium-high revenue
    for name in gold_names:
        revenue = random.randint(200000, 850000)
        region = random.choice(regions)
        days_ago = random.randint(365*2, 365*6)
        join_date = datetime.date(2025, 3, 4) - datetime.timedelta(days=days_ago)
        rows.append([f'C{cid:04d}', name, region, revenue, 'Gold', join_date.strftime('%Y-%m-%d')])
        cid += 1

    # Add Silver (~65 rows) - medium revenue
    for name in silver_names:
        revenue = random.randint(50000, 200000)
        region = random.choice(regions)
        days_ago = random.randint(365, 365*4)
        join_date = datetime.date(2025, 3, 4) - datetime.timedelta(days=days_ago)
        rows.append([f'C{cid:04d}', name, region, revenue, 'Silver', join_date.strftime('%Y-%m-%d')])
        cid += 1

    # Add Bronze (~30 rows) - lower revenue
    for name in bronze_names:
        revenue = random.randint(10000, 50000)
        region = random.choice(regions)
        days_ago = random.randint(180, 365*3)
        join_date = datetime.date(2025, 3, 4) - datetime.timedelta(days=days_ago)
        rows.append([f'C{cid:04d}', name, region, revenue, 'Bronze', join_date.strftime('%Y-%m-%d')])
        cid += 1

    # Fill remaining rows to reach 200 total (we have 20+45+65+30=160, need 40 more)
    extra_tiers = ['Silver'] * 25 + ['Bronze'] * 15
    extra_bases = [
        'Hillside Corp', 'Meadow Corp', 'Summit Corp', 'Valley Corp', 'Ridge Corp',
        'Creek Corp', 'Lake Corp', 'Forest Corp', 'Shore Corp', 'Bay Corp',
        'Peak Corp', 'Glen Corp', 'Cove Corp', 'Bluff Corp', 'Heights Corp',
        'Falls Corp', 'Springs Corp', 'Hollow Corp', 'Crossing Corp', 'Bend Corp',
        'Point Corp', 'Crest Corp', 'Knoll Corp', 'Cliff Corp', 'Ledge Corp',
        'Brook Corp', 'Stream Corp', 'Pond Corp', 'River Corp', 'Marsh Corp',
        'Trail Corp', 'Path Corp', 'Way Corp', 'Route Corp', 'Lane Corp',
        'Court Corp', 'Place Corp', 'Drive Corp', 'Avenue Corp', 'Circle Corp'
    ]
    for i, (t, nm) in enumerate(zip(extra_tiers, extra_bases)):
        revenue = random.randint(10000, 150000) if t == 'Silver' else random.randint(5000, 30000)
        region = random.choice(regions)
        days_ago = random.randint(90, 365*3)
        join_date = datetime.date(2025, 3, 4) - datetime.timedelta(days=days_ago)
        rows.append([f'C{cid:04d}', nm, region, revenue, t, join_date.strftime('%Y-%m-%d')])
        cid += 1

    # Shuffle to mix tiers (keep random.seed for reproducibility)
    random.shuffle(rows)

    for r, row_data in enumerate(rows, 2):
        for c, val in enumerate(row_data, 1):
            ws_cust.cell(row=r, column=c, value=val)

    # Column widths
    ws_cust.column_dimensions['A'].width = 10
    ws_cust.column_dimensions['B'].width = 28
    ws_cust.column_dimensions['C'].width = 16
    ws_cust.column_dimensions['D'].width = 14
    ws_cust.column_dimensions['E'].width = 12
    ws_cust.column_dimensions['F'].width = 12
    ws_cust.freeze_panes = 'A2'

    # ----------------------------------------------------------------
    # Sheet 2: CommRates - Commission Rate Lookup Table A1:B5
    # ----------------------------------------------------------------
    ws_comm = wb.create_sheet('CommRates')
    ws_comm['A1'] = 'Tier'
    ws_comm['B1'] = 'CommissionRate'
    ws_comm['A1'].font = Font(bold=True)
    ws_comm['B1'].font = Font(bold=True)

    comm_data = [
        ['Platinum', 0.08],
        ['Gold', 0.06],
        ['Silver', 0.04],
        ['Bronze', 0.02],
    ]
    for r, row_data in enumerate(comm_data, 2):
        ws_comm.cell(row=r, column=1, value=row_data[0])
        cell = ws_comm.cell(row=r, column=2, value=row_data[1])
        cell.number_format = '0%'

    ws_comm.column_dimensions['A'].width = 12
    ws_comm.column_dimensions['B'].width = 18

    # ----------------------------------------------------------------
    # Sheet 3: Summary - uses cell references (NOT named ranges yet)
    # ----------------------------------------------------------------
    ws_summ = wb.create_sheet('Summary')

    ws_summ['A1'] = 'Tier'
    ws_summ['B1'] = 'Total Revenue'
    ws_summ['A1'].font = Font(bold=True)
    ws_summ['B1'].font = Font(bold=True)

    ws_summ['A2'] = 'Platinum'
    ws_summ['A3'] = 'Gold'
    ws_summ['A4'] = 'Silver'
    ws_summ['A5'] = 'Bronze'

    # Formulas using cell references (no named ranges)
    ws_summ['B2'] = "=SUMIF(Customers.E:E,\"Platinum\",Customers.D:D)"
    ws_summ['B3'] = "=SUMIF(Customers.E:E,\"Gold\",Customers.D:D)"
    ws_summ['B4'] = "=SUMIF(Customers.E:E,\"Silver\",Customers.D:D)"
    ws_summ['B5'] = "=SUMIF(Customers.E:E,\"Bronze\",Customers.D:D)"

    ws_summ.column_dimensions['A'].width = 12
    ws_summ.column_dimensions['B'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Customers sheet: 200 rows of customer data (20 Platinum, 45 Gold, 65 Silver, 30+ Bronze)')
    print(f'  CommRates sheet: A1:B5 commission rate lookup table')
    print(f'  Summary sheet: formulas using direct cell references (no named ranges)')
    print(f'  No named ranges defined in workbook')


create_initial()
