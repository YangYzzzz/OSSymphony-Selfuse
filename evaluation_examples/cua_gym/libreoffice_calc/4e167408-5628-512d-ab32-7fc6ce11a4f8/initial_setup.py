"""
Initial Setup: Kitting Availability Check - BOM and Component Stock
Task ID: calc_ops_inventory_kitting_bom_069
Domain: libreoffice_calc

Creates a spreadsheet with:
- KitBOM sheet: Bill of Materials for Kit-A, Kit-B, Kit-C
- ComponentStock sheet: Current on-hand quantities
- KitPlanning sheet: Empty analysis table (formulas to be filled in by agent)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_inventory_kitting_bom_069'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # =========================================================
    # Sheet 1: KitBOM
    # =========================================================
    ws_bom = wb.active
    ws_bom.title = 'KitBOM'

    # Headers row 1
    bom_headers = [
        'Component SKU', 'Component Name',
        'Kit-A Qty Required', 'Kit-B Qty Required', 'Kit-C Qty Required'
    ]
    for col, h in enumerate(bom_headers, 1):
        cell = ws_bom.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # BOM data: 10 components (rows 2-11)
    # SKUs, component names, qty needed for Kit-A, Kit-B, Kit-C
    bom_data = [
        ('SKU-1001', 'Aluminum Frame Bracket',    4,  2,  6),
        ('SKU-1002', 'M5 Hex Bolt (pack)',        8,  4, 12),
        ('SKU-1003', 'Neoprene Gasket',           2,  2,  3),
        ('SKU-1004', 'Steel Bearing Assembly',    1,  1,  2),
        ('SKU-1005', 'Connector Cable 30cm',      3,  2,  4),
        ('SKU-1006', 'PCB Control Module',        1,  1,  1),
        ('SKU-1007', 'Rubber O-Ring Set',         5,  3,  8),
        ('SKU-1008', 'Zinc-Coated Spring',       10,  6, 15),
        ('SKU-1009', 'Heat Shrink Tubing (m)',    2,  1,  3),
        ('SKU-1010', 'ABS Enclosure Panel',       2,  2,  4),
    ]

    for row_idx, row_data in enumerate(bom_data, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws_bom.cell(row=row_idx, column=col_idx, value=val)

    # Column widths
    ws_bom.column_dimensions['A'].width = 14
    ws_bom.column_dimensions['B'].width = 28
    ws_bom.column_dimensions['C'].width = 20
    ws_bom.column_dimensions['D'].width = 20
    ws_bom.column_dimensions['E'].width = 20

    # =========================================================
    # Sheet 2: ComponentStock
    # =========================================================
    ws_stock = wb.create_sheet('ComponentStock')

    # Headers row 1
    stock_headers = ['SKU', 'On Hand Qty']
    for col, h in enumerate(stock_headers, 1):
        cell = ws_stock.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Stock data: matching SKUs with on-hand quantities
    # Designed so some components will be binding constraints
    stock_data = [
        ('SKU-1001', 320),
        ('SKU-1002', 580),
        ('SKU-1003', 210),
        ('SKU-1004',  85),
        ('SKU-1005', 275),
        ('SKU-1006',  72),
        ('SKU-1007', 430),
        ('SKU-1008', 820),
        ('SKU-1009', 190),
        ('SKU-1010', 195),
    ]

    for row_idx, row_data in enumerate(stock_data, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws_stock.cell(row=row_idx, column=col_idx, value=val)

    ws_stock.column_dimensions['A'].width = 14
    ws_stock.column_dimensions['B'].width = 16

    # =========================================================
    # Sheet 3: KitPlanning
    # =========================================================
    ws_plan = wb.create_sheet('KitPlanning')

    # Requested quantity inputs at B1:B3
    ws_plan.cell(row=1, column=1, value='Kit-A Requested Qty')
    ws_plan.cell(row=1, column=2, value=50)   # B1
    ws_plan.cell(row=2, column=1, value='Kit-B Requested Qty')
    ws_plan.cell(row=2, column=2, value=30)   # B2
    ws_plan.cell(row=3, column=1, value='Kit-C Requested Qty')
    ws_plan.cell(row=3, column=2, value=20)   # B3

    # Row 4 is blank spacer

    # Row 5 headers for component analysis table (rows 5-14 = 10 components)
    plan_headers = [
        'Component SKU', 'On Hand', 'Required Total',
        'Available for Kitting', 'Binding Constraint'
    ]
    for col, h in enumerate(plan_headers, 1):
        cell = ws_plan.cell(row=5, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Rows 6-15: Component SKUs pre-filled (A column only), B-E left empty for agent
    # NOTE: context says rows 5-14 are the "component analysis table headers"
    # and B-E are "empty" - agent must add formulas.
    # The SKUs in column A help agent know what to reference.
    for row_idx, sku in enumerate(
        ['SKU-1001','SKU-1002','SKU-1003','SKU-1004','SKU-1005',
         'SKU-1006','SKU-1007','SKU-1008','SKU-1009','SKU-1010'], 6
    ):
        ws_plan.cell(row=row_idx, column=1, value=sku)
        # B (On Hand), C (Required Total), D (Available), E (Binding) - all empty

    ws_plan.column_dimensions['A'].width = 16
    ws_plan.column_dimensions['B'].width = 14
    ws_plan.column_dimensions['C'].width = 18
    ws_plan.column_dimensions['D'].width = 22
    ws_plan.column_dimensions['E'].width = 20

    # =========================================================
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: KitBOM, ComponentStock, KitPlanning')
    print('  KitBOM: 10 components, 5 columns (SKU, Name, Kit-A/B/C qty)')
    print('  ComponentStock: 10 components with on-hand qty')
    print('  KitPlanning: requested qty inputs at B1:B3; analysis table rows 5-14 with SKUs in col A; B-E empty')


create_initial()
