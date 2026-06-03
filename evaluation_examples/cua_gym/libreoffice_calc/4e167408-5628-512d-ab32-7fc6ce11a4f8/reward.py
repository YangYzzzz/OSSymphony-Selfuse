"""
Reward Script: Kitting availability check with BOM
Task ID: calc_ops_inventory_kitting_bom_069
Domain: libreoffice_calc
Scoring:
  Component 1: VLOOKUP formulas in KitPlanning B6:B15 pulling On Hand stock  (0.30)
  Component 2: Required Total formulas in KitPlanning C6:C15                  (0.25)
  Component 3: Available for Kitting formulas in KitPlanning D6:D15           (0.20)
  Component 4: Binding Constraint IF formulas in KitPlanning E6:E15           (0.15)
  Component 5: Max Buildable Kits summary section rows 17-20                  (0.10)
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_inventory_kitting_bom_069'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    The task requires filling in the KitPlanning sheet:
    - B6:B15: VLOOKUP formulas pulling On Hand qty from ComponentStock
    - C6:C15: Required Total = KitBOM qty * requested kit quantities
    - D6:D15: Available for Kitting = B - C (surplus or shortfall)
    - E6:E15: Binding Constraint = IF(D<0, "CONSTRAINT", "")
    - Rows 17-20: Summary showing max buildable kits for Kit-A, Kit-B, Kit-C
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check KitPlanning sheet exists
    if 'KitPlanning' not in wb.sheetnames:
        print("FAIL: KitPlanning sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['KitPlanning']

    # Component 1: VLOOKUP formulas in B6:B15 (On Hand stock lookup) (0.30 points)
    # Must reference ComponentStock and use VLOOKUP
    try:
        vlookup_count = 0
        vlookup_details = []
        for row in range(6, 16):  # rows 6 to 15 inclusive (10 rows)
            cell_val = ws.cell(row=row, column=2).value  # column B
            if cell_val is not None and isinstance(cell_val, str):
                val_upper = cell_val.upper().replace(' ', '')
                # Must contain VLOOKUP and reference ComponentStock
                if 'VLOOKUP' in val_upper and 'COMPONENTSTOCK' in val_upper:
                    vlookup_count += 1
                    vlookup_details.append(f"B{row}: {cell_val}")
            else:
                vlookup_details.append(f"B{row}: {repr(cell_val)} (missing)")

        if vlookup_count == 10:
            print(f"PASS: Component 1 — All 10 VLOOKUP formulas found in B6:B15 (0.30 pts)")
            total_score += 0.30
        elif vlookup_count >= 5:
            partial = round(0.30 * vlookup_count / 10, 3)
            print(f"PARTIAL: Component 1 — {vlookup_count}/10 VLOOKUP formulas in B6:B15 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {vlookup_count}/10 VLOOKUP formulas in B6:B15")
            for d in vlookup_details[:3]:
                print(f"  {d}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Required Total formulas in C6:C15 (0.25 points)
    # Must reference KitBOM columns and $B$1, $B$2, $B$3 kit requested quantities
    try:
        req_formula_count = 0
        req_details = []
        for row in range(6, 16):  # rows 6 to 15 inclusive
            cell_val = ws.cell(row=row, column=3).value  # column C
            if cell_val is not None and isinstance(cell_val, str):
                val_upper = cell_val.upper().replace(' ', '')
                # Must reference KitBOM and use multiplication with $B$1/$B$2/$B$3
                if 'KITBOM' in val_upper and ('$B$1' in val_upper or '$B$2' in val_upper or '$B$3' in val_upper):
                    req_formula_count += 1
                    req_details.append(f"C{row}: {cell_val}")
            else:
                req_details.append(f"C{row}: {repr(cell_val)} (missing)")

        if req_formula_count == 10:
            print(f"PASS: Component 2 — All 10 Required Total formulas found in C6:C15 (0.25 pts)")
            total_score += 0.25
        elif req_formula_count >= 5:
            partial = round(0.25 * req_formula_count / 10, 3)
            print(f"PARTIAL: Component 2 — {req_formula_count}/10 Required Total formulas in C6:C15 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {req_formula_count}/10 Required Total formulas in C6:C15")
            for d in req_details[:3]:
                print(f"  {d}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Available for Kitting formulas in D6:D15 (0.20 points)
    # Must be =Bn-Cn (or equivalent subtraction referencing same row columns B and C)
    try:
        avail_formula_count = 0
        avail_details = []
        for row in range(6, 16):  # rows 6 to 15 inclusive
            cell_val = ws.cell(row=row, column=4).value  # column D
            if cell_val is not None and isinstance(cell_val, str):
                val_clean = cell_val.replace(' ', '')
                # Check for subtraction formula referencing B-C of same row
                pattern = re.compile(rf'=?B{row}-C{row}', re.IGNORECASE)
                if pattern.search(val_clean):
                    avail_formula_count += 1
                    avail_details.append(f"D{row}: {cell_val}")
                else:
                    avail_details.append(f"D{row}: {repr(cell_val)} (unexpected pattern)")
            else:
                avail_details.append(f"D{row}: {repr(cell_val)} (missing)")

        if avail_formula_count == 10:
            print(f"PASS: Component 3 — All 10 Available for Kitting formulas (B-C) found in D6:D15 (0.20 pts)")
            total_score += 0.20
        elif avail_formula_count >= 5:
            partial = round(0.20 * avail_formula_count / 10, 3)
            print(f"PARTIAL: Component 3 — {avail_formula_count}/10 B-C formulas in D6:D15 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {avail_formula_count}/10 formulas in D6:D15")
            for d in avail_details[:3]:
                print(f"  {d}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Binding Constraint IF formulas in E6:E15 (0.15 points)
    # Must contain IF(...<0...) referencing D column and return "CONSTRAINT"
    try:
        constraint_formula_count = 0
        constraint_details = []
        for row in range(6, 16):  # rows 6 to 15 inclusive
            cell_val = ws.cell(row=row, column=5).value  # column E
            if cell_val is not None and isinstance(cell_val, str):
                val_upper = cell_val.upper().replace(' ', '')
                # Must use IF, reference D column, and include "CONSTRAINT"
                if 'IF(' in val_upper and 'CONSTRAINT' in val_upper and f'D{row}' in cell_val:
                    constraint_formula_count += 1
                    constraint_details.append(f"E{row}: {cell_val}")
                else:
                    constraint_details.append(f"E{row}: {repr(cell_val)} (unexpected pattern)")
            else:
                constraint_details.append(f"E{row}: {repr(cell_val)} (missing)")

        if constraint_formula_count == 10:
            print(f"PASS: Component 4 — All 10 Binding Constraint IF formulas found in E6:E15 (0.15 pts)")
            total_score += 0.15
        elif constraint_formula_count >= 5:
            partial = round(0.15 * constraint_formula_count / 10, 3)
            print(f"PARTIAL: Component 4 — {constraint_formula_count}/10 IF formulas in E6:E15 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {constraint_formula_count}/10 IF formulas in E6:E15")
            for d in constraint_details[:3]:
                print(f"  {d}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Max Buildable Kits summary section (rows 17-20) (0.10 points)
    # Row 17: header "Max Buildable Kits" in A17
    # Rows 18-20: Kit-A, Kit-B, Kit-C labels in A18-A20 with FLOOR/MIN formulas in B18-B20
    try:
        summary_checks_passed = 0
        summary_checks_total = 4  # A17 header + 3 kit rows
        summary_details = []

        # Check A17 header
        a17 = ws.cell(row=17, column=1).value
        if a17 is not None and 'max' in str(a17).lower():
            summary_checks_passed += 1
            summary_details.append(f"A17: {repr(a17)} (header present)")
        else:
            summary_details.append(f"A17: {repr(a17)} (expected 'Max Buildable Kits')")

        # Check labels and formulas in rows 18-20 (Kit-A, Kit-B, Kit-C)
        kit_labels_expected = ['Kit-A', 'Kit-B', 'Kit-C']
        for row, kit in zip([18, 19, 20], kit_labels_expected):
            label = ws.cell(row=row, column=1).value
            formula = ws.cell(row=row, column=2).value
            label_ok = label is not None and kit.lower() in str(label).lower()
            formula_ok = (formula is not None and isinstance(formula, str) and
                         ('FLOOR' in formula.upper() or 'MIN' in formula.upper()) and
                         'VLOOKUP' in formula.upper())
            if label_ok and formula_ok:
                summary_checks_passed += 1
                summary_details.append(f"Row {row}: label={repr(label)}, formula present")
            elif label_ok:
                summary_details.append(f"Row {row}: label={repr(label)}, formula missing/incorrect: {repr(formula)}")
            else:
                summary_details.append(f"Row {row}: label={repr(label)}, formula={repr(formula)} (unexpected)")

        for d in summary_details:
            print(f"  {d}")
        if summary_checks_passed == summary_checks_total:
            print(f"PASS: Component 5 — Max Buildable Kits summary section complete (0.10 pts)")
            total_score += 0.10
        elif summary_checks_passed > 0:
            summary_score = round(0.10 * summary_checks_passed / summary_checks_total, 4)
            print(f"PARTIAL: Component 5 — {summary_checks_passed}/{summary_checks_total} summary checks passed ({summary_score} pts)")
            total_score += summary_score
        else:
            print(f"FAIL: Component 5 — Summary section missing")

    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
