"""
Initial Setup: Apply Japanese Yen currency format to values in B2:B4
Task ID: calc_lf_056
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_056'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'JapanSales'

    # Headers
    ws['A1'] = 'Item'
    ws['B1'] = 'Price (JPY)'

    # Data rows - plain numeric values, NO currency formatting
    ws['A2'] = 'Item A'
    ws['B2'] = 15800

    ws['A3'] = 'Item B'
    ws['B3'] = 3200

    ws['A4'] = 'Item C'
    ws['B4'] = 98500

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
