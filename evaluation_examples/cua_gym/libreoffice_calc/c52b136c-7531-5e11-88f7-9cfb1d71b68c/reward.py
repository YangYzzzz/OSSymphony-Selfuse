"""
Reward Script: Apply Japanese Yen currency format to B2:B4
Task ID: calc_lf_056
Domain: libreoffice_calc
Scoring:
  Component 1: B2 has Yen format '[$¥-411]#,##0' and correct value (0.34 pts)
  Component 2: B3 has Yen format '[$¥-411]#,##0' and correct value (0.33 pts)
  Component 3: B4 has Yen format '[$¥-411]#,##0' and correct value (0.33 pts)
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_056'
EXPECTED_FORMAT = '[$\u00a5-411]#,##0'

# Expected values for each cell (also verifies data integrity as sub-condition)
EXPECTED_VALUES = {
    'B2': 15800,
    'B3': 3200,
    'B4': 98500,
}

WEIGHTS = {
    'B2': 0.34,
    'B3': 0.33,
    'B4': 0.33,
}


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet exists
    if 'JapanSales' not in wb.sheetnames:
        print("FAIL: Sheet 'JapanSales' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['JapanSales']

    for cell_ref in ['B2', 'B3', 'B4']:
        weight = WEIGHTS[cell_ref]
        expected_val = EXPECTED_VALUES[cell_ref]

        # Component: <cell_ref> has Yen format and correct value
        try:
            cell = ws[cell_ref]
            cell_val = cell.value
            cell_fmt = cell.number_format

            # Check that the number format matches the expected Yen format
            if cell_fmt == EXPECTED_FORMAT and cell_val == expected_val:
                print(f"PASS: {cell_ref} — format='{cell_fmt}', value={cell_val} ({weight} pts)")
                total_score += weight
            elif cell_fmt == EXPECTED_FORMAT and cell_val != expected_val:
                print(f"FAIL: {cell_ref} — format is correct but value={cell_val}, expected {expected_val}")
            elif cell_fmt != EXPECTED_FORMAT and cell_val == expected_val:
                print(f"FAIL: {cell_ref} — value is correct but format='{cell_fmt}', expected '{EXPECTED_FORMAT}'")
            else:
                print(f"FAIL: {cell_ref} — format='{cell_fmt}' (expected '{EXPECTED_FORMAT}'), value={cell_val} (expected {expected_val})")
        except Exception as e:
            print(f"ERROR: {cell_ref} — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
