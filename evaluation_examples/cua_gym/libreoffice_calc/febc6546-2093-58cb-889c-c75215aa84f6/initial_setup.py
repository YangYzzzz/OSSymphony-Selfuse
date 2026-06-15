"""
Initial Setup: Create inventory spreadsheet with 160 rows of product data
Task ID: calc_pivot_015
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_015'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventory'

    # --- Headers ---
    headers = ['SKU', 'ProductName', 'Warehouse', 'Quantity', 'UnitCost', 'LastRestocked']
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Warehouse distribution ---
    # Target sums: WH-North=2400, WH-South=1850, WH-East=3100, WH-West=1600, WH-Central=2750
    # Grand total=11700
    # We assign 32 rows per warehouse (32*5=160)
    warehouses_config = [
        ('WH-North', 32, 2400),
        ('WH-South', 32, 1850),
        ('WH-East', 32, 3100),
        ('WH-West', 32, 1600),
        ('WH-Central', 32, 2750),
    ]

    # Product name components
    adjectives = ['Premium', 'Standard', 'Economy', 'Deluxe', 'Pro', 'Ultra', 'Compact', 'Heavy-Duty',
                  'Industrial', 'Commercial', 'Portable', 'Advanced', 'Classic', 'Elite', 'Basic']
    products = ['Widget', 'Sprocket', 'Bearing', 'Gasket', 'Valve', 'Motor', 'Sensor', 'Filter',
                'Pump', 'Bracket', 'Connector', 'Switch', 'Relay', 'Actuator', 'Coupling',
                'Fastener', 'Bushing', 'Flange', 'Adapter', 'Regulator', 'Capacitor', 'Resistor',
                'Transformer', 'Insulator', 'Conductor']

    # Dates for LastRestocked (2024-2025 range)
    months_days = []
    for y in [2024, 2025]:
        for m in range(1, 13):
            for d in [5, 12, 18, 25]:
                months_days.append(f'{y}-{m:02d}-{d:02d}')

    row_idx = 2
    sku_num = 1

    for wh_name, count, target_sum in warehouses_config:
        # Generate random quantities that sum to target_sum
        # Start with base quantities, then adjust last one
        quantities = []
        remaining = target_sum
        for i in range(count):
            if i == count - 1:
                quantities.append(remaining)
            else:
                # Random quantity between 30 and 150, but leave room
                avg_remaining = remaining / (count - i)
                low = max(10, int(avg_remaining * 0.4))
                high = min(200, int(avg_remaining * 1.6))
                q = random.randint(low, high)
                # Ensure we don't overshoot
                max_allowed = remaining - (count - i - 1) * 10
                q = min(q, max_allowed)
                quantities.append(q)
                remaining -= q

        random.shuffle(quantities)

        for i in range(count):
            adj = random.choice(adjectives)
            prod = random.choice(products)
            product_name = f'{adj} {prod}'
            quantity = quantities[i]
            unit_cost = round(random.uniform(2.50, 85.00), 2)
            date_str = random.choice(months_days)

            ws.cell(row=row_idx, column=1, value=sku_num)
            ws.cell(row=row_idx, column=2, value=product_name)
            ws.cell(row=row_idx, column=3, value=wh_name)
            ws.cell(row=row_idx, column=4, value=quantity)
            ws.cell(row=row_idx, column=5, value=unit_cost)
            ws.cell(row=row_idx, column=5).number_format = '$#,##0.00'
            ws.cell(row=row_idx, column=6, value=date_str)

            row_idx += 1
            sku_num += 1

    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 16

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Verify sums
    wb2 = openpyxl.load_workbook(OUTPUT)
    ws2 = wb2['Inventory']
    sums = {}
    for r in range(2, 162):
        wh = ws2.cell(row=r, column=3).value
        qty = ws2.cell(row=r, column=4).value
        sums[wh] = sums.get(wh, 0) + qty
    print(f'Warehouse sums: {sums}')
    print(f'Grand total: {sum(sums.values())}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
