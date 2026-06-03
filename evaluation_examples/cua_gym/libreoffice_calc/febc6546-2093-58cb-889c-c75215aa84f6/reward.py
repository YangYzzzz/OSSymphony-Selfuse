"""
Reward Script: Create a pivot table from inventory data showing total stock quantity per warehouse location.
Task ID: calc_pivot_015
Domain: libreoffice_calc
Scoring:
  Component 1: PivotTable sheet exists (0.15 pts)
  Component 2: Warehouse labels present in pivot table (0.20 pts)
  Component 3: Sum of Quantity values match expected per warehouse (0.40 pts)
  Component 4: Grand Total row present and correct (0.25 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_015'

# Expected ground truth values from task context
EXPECTED_WAREHOUSE_TOTALS = {
    'WH-North': 2400,
    'WH-South': 1850,
    'WH-East': 3100,
    'WH-West': 1600,
    'WH-Central': 2750,
}
EXPECTED_GRAND_TOTAL = 11700


def find_pivot_sheet(wb):
    """Find a sheet that looks like a pivot table (not the Inventory sheet).
    Returns the worksheet or None."""
    for name in wb.sheetnames:
        if name.lower() == 'inventory':
            continue
        ws = wb[name]
        # Check if this sheet has warehouse-like data
        # A pivot table sheet should have relatively few rows (< 20)
        if ws.max_row is not None and ws.max_row <= 30:
            return ws
    return None


def extract_pivot_data(ws):
    """Extract warehouse -> quantity mapping and grand total from pivot sheet.
    Returns (dict of warehouse: quantity, grand_total or None)."""
    warehouse_totals = {}
    grand_total = None

    # Scan all rows looking for warehouse names and numeric values
    for row in range(1, (ws.max_row or 0) + 1):
        # Check columns A and B (typical pivot layout)
        for label_col, val_col in [(1, 2), (2, 3), (1, 3)]:
            label = ws.cell(row=row, column=label_col).value
            val = ws.cell(row=row, column=val_col).value

            if label is None or val is None:
                continue

            label_str = str(label).strip()

            # Check if this is a warehouse row
            for wh_name in EXPECTED_WAREHOUSE_TOTALS:
                if wh_name.lower() == label_str.lower():
                    try:
                        warehouse_totals[wh_name] = float(val)
                    except (ValueError, TypeError):
                        pass
                    break

            # Check for grand total row
            if 'grand' in label_str.lower() and 'total' in label_str.lower():
                try:
                    grand_total = float(val)
                except (ValueError, TypeError):
                    pass
            elif label_str.lower() == 'total':
                try:
                    grand_total = float(val)
                except (ValueError, TypeError):
                    pass

    return warehouse_totals, grand_total


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: A separate pivot table sheet exists (0.15 points)
    # This checks that a new sheet was created beyond the original 'Inventory' sheet.
    try:
        pivot_ws = find_pivot_sheet(wb)
        if pivot_ws is not None:
            print(f"PASS: Component 1 -- Pivot table sheet found: '{pivot_ws.title}' (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 -- No pivot table sheet found. Sheets: {wb.sheetnames}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    if pivot_ws is None:
        # No pivot sheet means nothing else to check
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    # Extract pivot data
    warehouse_totals, grand_total = extract_pivot_data(pivot_ws)
    print(f"  Extracted warehouse totals: {warehouse_totals}")
    print(f"  Extracted grand total: {grand_total}")

    # Component 2: Warehouse labels present in pivot table (0.20 points)
    # Check that all 5 warehouse names appear as row labels.
    try:
        found_warehouses = set(warehouse_totals.keys())
        expected_warehouses = set(EXPECTED_WAREHOUSE_TOTALS.keys())
        if found_warehouses == expected_warehouses:
            print(f"PASS: Component 2 -- All 5 warehouse labels found (0.20 pts)")
            total_score += 0.20
        else:
            missing = expected_warehouses - found_warehouses
            extra = found_warehouses - expected_warehouses
            print(f"FAIL: Component 2 -- Missing: {missing}, Extra: {extra}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Sum of Quantity values match expected per warehouse (0.40 points)
    # Each warehouse match earns 0.08 points (5 * 0.08 = 0.40).
    try:
        correct_count = 0
        for wh_name, expected_val in EXPECTED_WAREHOUSE_TOTALS.items():
            actual_val = warehouse_totals.get(wh_name)
            if actual_val is not None and abs(actual_val - expected_val) < 1.0:
                print(f"  PASS: {wh_name} = {actual_val} (expected {expected_val})")
                correct_count += 1
            else:
                print(f"  FAIL: {wh_name} = {actual_val} (expected {expected_val})")

        wh_score = correct_count * 0.08
        if correct_count == len(EXPECTED_WAREHOUSE_TOTALS):
            print(f"PASS: Component 3 -- All 5 warehouse quantities correct (0.40 pts)")
            total_score += wh_score
        elif correct_count > 0:
            print(f"PARTIAL: Component 3 -- {correct_count}/5 correct ({wh_score:.2f} pts)")
            total_score += wh_score
        else:
            print(f"FAIL: Component 3 -- No warehouse quantities matched")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Grand Total row present and correct (0.25 points)
    # The pivot table should include a grand total of 11700.
    try:
        if grand_total is not None and abs(grand_total - EXPECTED_GRAND_TOTAL) < 1.0:
            print(f"PASS: Component 4 -- Grand total = {grand_total} (expected {EXPECTED_GRAND_TOTAL}) (0.25 pts)")
            total_score += 0.25
        elif grand_total is not None:
            print(f"FAIL: Component 4 -- Grand total = {grand_total} (expected {EXPECTED_GRAND_TOTAL})")
        else:
            print(f"FAIL: Component 4 -- No grand total row found")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved LibreOffice edits before verification
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
persist_app_state()

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
