"""
Initial Setup: Remove and re-protect sheet with new password
Task ID: calc_ps_034
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.worksheet.protection import SheetProtection

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_034'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Archive sheet ---
    ws = wb.active
    ws.title = 'Archive'

    # Headers
    headers = ['Employee ID', 'Full Name', 'Department', 'Hire Date', 'Annual Salary', 'Performance Rating']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic employee data for 199 rows (rows 2-200 => A1:F200 total)
    departments = ['Engineering', 'Marketing', 'Finance', 'HR', 'Operations', 'Sales', 'Legal', 'R&D']
    first_names = [
        'Sarah', 'Marcus', 'Elena', 'James', 'Priya', 'Carlos', 'Aisha', 'David',
        'Mei', 'Robert', 'Fatima', 'Thomas', 'Yuki', 'Michael', 'Olga',
        'Samuel', 'Lina', 'Patrick', 'Zara', 'William', 'Nadia', 'Henry',
        'Suki', 'George', 'Amara', 'Victor', 'Rosa', 'Ivan', 'Keiko', 'Leo'
    ]
    last_names = [
        'Chen', 'Johnson', 'Petrov', 'Williams', 'Sharma', 'Rodriguez', 'Hassan', 'Kim',
        'Liu', 'Brown', 'Ali', 'Mueller', 'Tanaka', 'Davis', 'Ivanova',
        'Okafor', 'Johansson', 'Reilly', 'Patel', 'Thompson', 'Khoury', 'Park',
        'Nakamura', 'Wilson', 'Diallo', 'Santos', 'Garcia', 'Novak', 'Yamamoto', 'Fischer'
    ]
    ratings = ['Excellent', 'Good', 'Satisfactory', 'Needs Improvement', 'Outstanding']

    random.seed(42)
    for r in range(2, 201):
        emp_id = f'EMP-{1000 + r - 1:04d}'
        fname = random.choice(first_names)
        lname = random.choice(last_names)
        dept = random.choice(departments)
        year = random.randint(2015, 2025)
        month = random.randint(1, 12)
        day = random.randint(1, 28)
        hire_date = f'{year}-{month:02d}-{day:02d}'
        salary = random.randint(45000, 165000)
        rating = random.choice(ratings)

        ws.cell(row=r, column=1, value=emp_id)
        ws.cell(row=r, column=2, value=f'{fname} {lname}')
        ws.cell(row=r, column=3, value=dept)
        ws.cell(row=r, column=4, value=hire_date)
        ws.cell(row=r, column=5, value=salary)
        ws.cell(row=r, column=6, value=rating)

    # Protect the Archive sheet with password 'arch2023', allowing sort
    ws.protection = SheetProtection(
        sheet=True,
        password='arch2023',
        sort=False,           # False means sorting IS allowed
        autoFilter=False,     # Also allow auto filter for sorting support
    )

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
