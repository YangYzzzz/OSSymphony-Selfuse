"""
Reward Script: Balanced Scorecard Dashboard
Task ID: calc_gsd_045
Domain: libreoffice_calc
Scoring:
  Component 1 (0.15): Title row A1:G1 merged with correct text and formatting
  Component 2 (0.40): Four section headers with correct text, merge, bg color, white bold text
  Component 3 (0.20): KPI data rows populated in each section (4 KPIs x 4 sections)
  Component 4 (0.10): Conditional formatting on Status columns (E4:E7, E10:E13, E16:E19, E22:E25)
  Component 5 (0.15): Summary section at rows 27-31 with data
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_045'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Dashboard sheet must exist
    if 'Dashboard' not in wb.sheetnames:
        print("FAIL: 'Dashboard' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Dashboard']

    # =========================================================================
    # Component 1: Title row A1:G1 merged with correct text and formatting (0.15 pts)
    # =========================================================================
    try:
        # Check A1:G1 is merged
        title_merged = False
        for mr in ws.merged_cells.ranges:
            if str(mr) == 'A1:G1':
                title_merged = True
                break

        a1 = ws['A1']
        title_text = str(a1.value).strip() if a1.value else ''
        title_correct_text = 'Balanced Scorecard Q4 2024' in title_text
        title_bold = a1.font.bold is True
        title_size = a1.font.size is not None and a1.font.size >= 14  # allow some tolerance

        if title_merged and title_correct_text and title_bold and title_size:
            print(f"PASS: Component 1 - Title merged A1:G1, text='{title_text}', bold={title_bold}, size={a1.font.size} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 - merged={title_merged}, text_match={title_correct_text}, bold={title_bold}, size={a1.font.size}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # =========================================================================
    # Component 2: Four section headers with correct properties (0.40 pts, 0.10 each)
    # =========================================================================
    section_checks = [
        ('A3', 'A3:G3', 'FINANCIAL PERSPECTIVE', 'FF006400'),   # dark green
        ('A9', 'A9:G9', 'CUSTOMER PERSPECTIVE', 'FF00008B'),     # dark blue
        ('A15', 'A15:G15', 'INTERNAL PROCESS', 'FF800080'),      # purple
        ('A21', 'A21:G21', 'LEARNING & GROWTH', 'FFCC5500'),     # dark orange
    ]

    for cell_addr, merge_range, expected_text, expected_fill in section_checks:
        try:
            cell = ws[cell_addr]
            cell_val = str(cell.value).strip() if cell.value else ''

            # Check merge
            is_merged = False
            for mr in ws.merged_cells.ranges:
                if str(mr) == merge_range:
                    is_merged = True
                    break

            # Check text
            text_ok = expected_text in cell_val.upper()

            # Check background fill color
            fill_rgb = None
            try:
                fill_rgb = cell.fill.fgColor.rgb
            except:
                pass
            fill_ok = fill_rgb == expected_fill

            # Check white font
            font_white = False
            try:
                font_rgb = cell.font.color.rgb
                # White can be 00FFFFFF or FFFFFFFF
                font_white = font_rgb is not None and 'FFFFFF' in font_rgb
            except:
                pass

            # Check bold
            bold_ok = cell.font.bold is True

            if is_merged and text_ok and fill_ok and font_white and bold_ok:
                print(f"PASS: Component 2 - {cell_addr} header correct: '{cell_val}', fill={fill_rgb}, white text, bold (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 2 - {cell_addr}: merged={is_merged}, text_ok={text_ok}('{cell_val}'), fill={fill_rgb}(expect {expected_fill}), white={font_white}, bold={bold_ok}")
        except Exception as e:
            print(f"ERROR: Component 2 - {cell_addr}: {e}")

    # =========================================================================
    # Component 3: KPI data rows populated (0.20 pts, 0.05 per section)
    # =========================================================================
    kpi_sections = [
        ('Financial', [(4, 'Revenue Growth'), (5, 'Profit Margin'), (6, 'ROE'), (7, 'Cost Reduction')]),
        ('Customer', [(10, 'Customer'), (11, 'Customer'), (12, 'Net Promoter'), (13, 'Market Share')]),
        ('Internal', [(16, 'Process'), (17, 'Quality'), (18, 'Cycle'), (19, 'Innovation')]),
        ('Learning', [(22, 'Employee'), (23, 'Training'), (24, 'Employee'), (25, 'Skills')]),
    ]

    for section_name, kpi_rows in kpi_sections:
        try:
            rows_with_data = 0
            for row_num, kpi_keyword in kpi_rows:
                a_val = ws.cell(row=row_num, column=1).value
                if a_val and kpi_keyword.lower() in str(a_val).lower():
                    # Check that at least Target (B) and Actual (C) have numeric values
                    b_val = ws.cell(row=row_num, column=2).value
                    c_val = ws.cell(row=row_num, column=3).value
                    if b_val is not None and c_val is not None:
                        rows_with_data += 1

            if rows_with_data >= 3:  # at least 3 of 4 KPIs populated
                print(f"PASS: Component 3 - {section_name} has {rows_with_data}/4 KPI rows with data (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 3 - {section_name} only {rows_with_data}/4 KPI rows have data")
        except Exception as e:
            print(f"ERROR: Component 3 - {section_name}: {e}")

    # =========================================================================
    # Component 4: Conditional formatting on Status columns (0.10 pts)
    # =========================================================================
    try:
        cf_ranges = set()
        for cf in ws.conditional_formatting:
            cf_str = str(cf)
            # Normalize: extract the cell range portion
            cf_ranges.add(cf_str.replace('<ConditionalFormatting ', '').replace('>', '').strip())

        expected_cf = ['E4:E7', 'E10:E13', 'E16:E19', 'E22:E25']
        cf_found = 0
        for ecf in expected_cf:
            if ecf in cf_ranges:
                cf_found += 1

        if cf_found >= 3:  # at least 3 of 4 sections have conditional formatting
            print(f"PASS: Component 4 - Conditional formatting found on {cf_found}/4 status ranges (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 - Only {cf_found}/4 status ranges have conditional formatting. Found ranges: {cf_ranges}")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # =========================================================================
    # Component 5: Summary section at rows 27-31 (0.15 pts)
    # =========================================================================
    try:
        # Check A27 merged header
        summary_merged = False
        for mr in ws.merged_cells.ranges:
            if str(mr) == 'A27:G27':
                summary_merged = True
                break

        a27_val = ws['A27'].value
        summary_header = a27_val is not None and 'SUMMARY' in str(a27_val).upper()

        # Check that rows 28-31 have data (at least perspective names and scores)
        summary_rows_with_data = 0
        for r in range(28, 33):  # check rows 28-32 to be safe
            a_val = ws.cell(row=r, column=1).value
            if a_val is not None and str(a_val).strip():
                summary_rows_with_data += 1

        if summary_merged and summary_header and summary_rows_with_data >= 3:
            print(f"PASS: Component 5 - Summary section: merged={summary_merged}, header='{a27_val}', data rows={summary_rows_with_data} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 5 - merged={summary_merged}, header={summary_header}('{a27_val}'), data_rows={summary_rows_with_data}")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
