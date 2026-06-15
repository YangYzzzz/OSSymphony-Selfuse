"""
Initial Setup: Reference Resolution Pipeline - papers_to_process folder with 5 ML PDFs
Task ID: osworld_multi_apps_web_references_015
Domain: libreoffice_calc (multi-app: Chrome, Writer, file system)
"""

import os
import shlex
import subprocess
import time

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_web_references_015'
DESKTOP = f'{WORKDIR}/Desktop'
PAPERS_DIR = f'{DESKTOP}/papers_to_process'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_pdfs():
    """Create 5 realistic ML paper PDFs with References sections."""
    try:
        from fpdf import FPDF
    except ImportError:
        subprocess.run(['pip3', 'install', 'fpdf2'], check=True)
        from fpdf import FPDF

    os.makedirs(PAPERS_DIR, exist_ok=True)

    # Paper 1: Attention Is All You Need style paper
    paper1 = {
        'filename': 'attention_transformers_2023.pdf',
        'title': 'Enhanced Transformer Architectures for Long-Range Dependencies',
        'authors': 'Zhang, W., Liu, H., Chen, M., & Park, S.',
        'abstract': (
            'We propose a novel transformer architecture that improves upon the standard '
            'self-attention mechanism by introducing sparse attention patterns combined '
            'with hierarchical position encodings. Our model achieves state-of-the-art '
            'performance on language modeling benchmarks while reducing computational '
            'complexity from O(n^2) to O(n log n).'
        ),
        'references': [
            {
                'num': 1,
                'text': 'Vaswani, A., Shazeer, N., Parmar, N., Uszkoreit, J., Jones, L., Gomez, A. N., Kaiser, L., & Polosukhin, I. (2017). Attention is all you need. Advances in Neural Information Processing Systems, 30, 5998-6008.',
                'doi': '10.48550/arXiv.1706.03762',
                'year': 2017
            },
            {
                'num': 2,
                'text': 'Devlin, J., Chang, M. W., Lee, K., & Toutanova, K. (2019). BERT: Pre-training of deep bidirectional transformers for language understanding. Proceedings of NAACL-HLT 2019, 4171-4186.',
                'doi': '10.18653/v1/N19-1423',
                'year': 2019
            },
            {
                'num': 3,
                'text': 'Brown, T., Mann, B., Ryder, N., Subbiah, M., Kaplan, J. D., Dhariwal, P., ... & Amodei, D. (2020). Language models are few-shot learners. Advances in Neural Information Processing Systems, 33, 1877-1901.',
                'doi': '10.48550/arXiv.2005.14165',
                'year': 2020
            },
            {
                'num': 4,
                'text': 'Kitaev, N., Kaiser, L., & Levskaya, A. (2020). Reformer: The efficient transformer. International Conference on Learning Representations.',
                'doi': '10.48550/arXiv.2001.04451',
                'year': 2020
            },
            {
                'num': 5,
                'text': 'Liu, Y., Ott, M., Goyal, N., Du, J., Joshi, M., Chen, D., ... & Stoyanov, V. (2019). RoBERTa: A robustly optimized BERT pretraining approach. arXiv preprint arXiv:1907.11692.',
                'doi': '10.48550/arXiv.1907.11692',
                'year': 2019
            },
            {
                'num': 6,
                'text': 'Radford, A., Wu, J., Child, R., Luan, D., Amodei, D., & Sutskever, I. (2019). Language models are unsupervised multitask learners. OpenAI Blog, 1(8), 9.',
                'doi': None,
                'year': 2019
            },
            {
                'num': 7,
                'text': 'Child, R., Gray, S., Radford, A., & Sutskever, I. (2019). Generating long sequences with sparse transformers. arXiv preprint arXiv:1904.10509.',
                'doi': '10.48550/arXiv.1904.10509',
                'year': 2019
            },
            {
                'num': 8,
                'text': 'Beltagy, I., Peters, M. E., & Cohan, A. (2020). Longformer: The long-document transformer. arXiv preprint arXiv:2004.05150.',
                'doi': '10.48550/arXiv.2004.05150',
                'year': 2020
            },
        ]
    }

    # Paper 2: Computer Vision / CNN paper
    paper2 = {
        'filename': 'vision_contrastive_2022.pdf',
        'title': 'Contrastive Visual Representation Learning with Multi-Scale Feature Alignment',
        'authors': 'Patel, R., Kim, J., Thompson, A., & Martinez, C.',
        'abstract': (
            'Self-supervised visual representation learning has made remarkable progress '
            'through contrastive objectives. We introduce a multi-scale feature alignment '
            'strategy that enables models to capture both local texture patterns and global '
            'semantic structures simultaneously. Our method improves linear probe accuracy '
            'on ImageNet by 3.2% over strong baselines.'
        ),
        'references': [
            {
                'num': 1,
                'text': 'He, K., Fan, H., Wu, Y., Xie, S., & Girshick, R. (2020). Momentum contrast for unsupervised visual representation learning. Proceedings of CVPR 2020, 9729-9738.',
                'doi': '10.1109/CVPR42600.2020.00975',
                'year': 2020
            },
            {
                'num': 2,
                'text': 'Chen, T., Kornblith, S., Norouzi, M., & Hinton, G. (2020). A simple framework for contrastive learning of visual representations. Proceedings of ICML 2020, 1597-1607.',
                'doi': '10.48550/arXiv.2002.05709',
                'year': 2020
            },
            {
                'num': 3,
                'text': 'Grill, J. B., Strub, F., Altche, F., Tallec, C., Richemond, P., Buchatskaya, E., ... & Valko, M. (2020). Bootstrap your own latent-a new approach to self-supervised learning. Advances in NeurIPS 33.',
                'doi': '10.48550/arXiv.2006.07733',
                'year': 2020
            },
            {
                'num': 4,
                'text': 'He, K., Zhang, X., Ren, S., & Sun, J. (2016). Deep residual learning for image recognition. Proceedings of CVPR 2016, 770-778.',
                'doi': '10.1109/CVPR.2016.90',
                'year': 2016
            },
            {
                'num': 5,
                'text': 'Dosovitskiy, A., Beyer, L., Kolesnikov, A., Weissenborn, D., Zhai, X., Unterthiner, T., ... & Houlsby, N. (2021). An image is worth 16x16 words: Transformers for image recognition at scale. ICLR 2021.',
                'doi': '10.48550/arXiv.2010.11929',
                'year': 2021
            },
            {
                'num': 6,
                'text': 'Caron, M., Misra, I., Mairal, J., Goyal, P., Bojanowski, P., & Joulin, A. (2020). Unsupervised learning of visual features by contrasting cluster assignments. NeurIPS 2020.',
                'doi': '10.48550/arXiv.2006.09882',
                'year': 2020
            },
            {
                'num': 7,
                'text': 'Bardes, A., Ponce, J., & LeCun, Y. (2022). VICReg: Variance-invariance-covariance regularization for self-supervised learning. ICLR 2022.',
                'doi': '10.48550/arXiv.2105.04906',
                'year': 2022
            },
            {
                'num': 8,
                'text': 'Zbontar, J., Jing, L., Misra, I., LeCun, Y., & Deny, S. (2021). Barlow twins: Self-supervised learning via redundancy reduction. ICML 2021, 12310-12320.',
                'doi': '10.48550/arXiv.2103.03230',
                'year': 2021
            },
            {
                'num': 9,
                'text': 'Chen, X., & He, K. (2021). Exploring simple siamese representation learning. CVPR 2021, 15750-15758.',
                'doi': '10.1109/CVPR46437.2021.01549',
                'year': 2021
            },
        ]
    }

    # Paper 3: Reinforcement Learning paper
    paper3 = {
        'filename': 'rl_policy_optimization_2023.pdf',
        'title': 'Sample-Efficient Policy Optimization via Hierarchical Intrinsic Rewards',
        'authors': 'Nakamura, T., Williams, B., Singh, P., & Rossi, F.',
        'abstract': (
            'We address the challenge of exploration in sparse-reward reinforcement learning '
            'environments by introducing a hierarchical intrinsic reward framework. Our approach '
            'combines count-based exploration bonuses at the state level with goal-conditioned '
            'subpolicies at the option level. Experiments on MiniGrid and Atari demonstrate '
            'significant improvements in sample efficiency over PPO and SAC baselines.'
        ),
        'references': [
            {
                'num': 1,
                'text': 'Schulman, J., Wolski, F., Dhariwal, P., Radford, A., & Klimov, O. (2017). Proximal policy optimization algorithms. arXiv preprint arXiv:1707.06347.',
                'doi': '10.48550/arXiv.1707.06347',
                'year': 2017
            },
            {
                'num': 2,
                'text': 'Mnih, V., Kavukcuoglu, K., Silver, D., Rusu, A. A., Veness, J., Bellemare, M. G., ... & Hassabis, D. (2015). Human-level control through deep reinforcement learning. Nature, 518(7540), 529-533.',
                'doi': '10.1038/nature14236',
                'year': 2015
            },
            {
                'num': 3,
                'text': 'Haarnoja, T., Zhou, A., Abbeel, P., & Levine, S. (2018). Soft actor-critic: Off-policy maximum entropy deep reinforcement learning with a stochastic actor. ICML 2018.',
                'doi': '10.48550/arXiv.1801.01290',
                'year': 2018
            },
            {
                'num': 4,
                'text': 'Pathak, D., Agrawal, P., Efros, A. A., & Darrell, T. (2017). Curiosity-driven exploration by self-supervised prediction. ICML 2017.',
                'doi': '10.1109/CVPRW.2017.70',
                'year': 2017
            },
            {
                'num': 5,
                'text': 'Barto, A. G., & Mahadevan, S. (2003). Recent advances in hierarchical reinforcement learning. Discrete Event Dynamic Systems, 13(1-2), 41-77.',
                'doi': '10.1023/A:1022140919877',
                'year': 2003
            },
            {
                'num': 6,
                'text': 'Nachum, O., Gu, S. S., Lee, H., & Levine, S. (2018). Data-efficient hierarchical reinforcement learning. NeurIPS 2018.',
                'doi': '10.48550/arXiv.1805.08296',
                'year': 2018
            },
            {
                'num': 7,
                'text': 'Bellemare, M. G., Srinivasan, S., Ostrovski, G., Schaul, T., Saxton, D., & Munos, R. (2016). Unifying count-based exploration and intrinsic motivation. NeurIPS 2016.',
                'doi': '10.48550/arXiv.1606.01868',
                'year': 2016
            },
            {
                'num': 8,
                'text': 'Chevalier-Boisvert, M., Willems, L., & Pal, S. (2018). Minimalistic gridworld environment for OpenAI gym. arXiv preprint arXiv:1801.01290.',
                'doi': None,
                'year': 2018
            },
        ]
    }

    # Paper 4: Graph Neural Networks paper
    paper4 = {
        'filename': 'graph_neural_nets_2022.pdf',
        'title': 'Scalable Graph Neural Networks with Dynamic Aggregation for Molecular Property Prediction',
        'authors': 'Anderson, L., Gupta, R., Chen, Y., & Muller, K.',
        'abstract': (
            'Molecular property prediction is a fundamental task in drug discovery and material '
            'science. We present DynAGG, a scalable graph neural network that employs dynamic '
            'aggregation functions conditioned on local chemical environment. DynAGG achieves '
            'new state-of-the-art results on QM9 and ZINC benchmarks while maintaining linear '
            'computational complexity with respect to graph size.'
        ),
        'references': [
            {
                'num': 1,
                'text': 'Kipf, T. N., & Welling, M. (2017). Semi-supervised classification with graph convolutional networks. ICLR 2017.',
                'doi': '10.48550/arXiv.1609.02907',
                'year': 2017
            },
            {
                'num': 2,
                'text': 'Xu, K., Hu, W., Leskovec, J., & Jegelka, S. (2019). How powerful are graph neural networks? ICLR 2019.',
                'doi': '10.48550/arXiv.1810.00826',
                'year': 2019
            },
            {
                'num': 3,
                'text': 'Gilmer, J., Schütt, K. T., & Tkatchenko, A. (2017). Neural message passing for quantum chemistry. ICML 2017, 1263-1272.',
                'doi': '10.48550/arXiv.1704.01212',
                'year': 2017
            },
            {
                'num': 4,
                'text': 'Velickovic, P., Cucurull, G., Casanova, A., Romero, A., Lio, P., & Bengio, Y. (2018). Graph attention networks. ICLR 2018.',
                'doi': '10.48550/arXiv.1710.10903',
                'year': 2018
            },
            {
                'num': 5,
                'text': 'Hamilton, W. L., Ying, R., & Leskovec, J. (2017). Inductive representation learning on large graphs. NeurIPS 2017.',
                'doi': '10.48550/arXiv.1706.02216',
                'year': 2017
            },
            {
                'num': 6,
                'text': 'Hu, W., Liu, B., Gomes, J., Zitnik, M., Liang, P., Pande, V., & Leskovec, J. (2020). Strategies for pre-training graph neural networks. ICLR 2020.',
                'doi': '10.48550/arXiv.1905.12265',
                'year': 2020
            },
            {
                'num': 7,
                'text': 'Ramakrishnan, R., Dral, P. O., Rupp, M., & von Lilienfeld, O. A. (2014). Quantum chemistry structures and properties of 134 kilo molecules. Scientific Data, 1, 140022.',
                'doi': '10.1038/sdata.2014.22',
                'year': 2014
            },
            {
                'num': 8,
                'text': 'Wieder, O., Kohlbacher, S., Kuenemann, M., Garon, A., Ducrot, P., Seidel, T., & Langer, T. (2020). A compact review of molecular property prediction with graph neural networks. Drug Discovery Today: Technologies, 37, 1-12.',
                'doi': '10.1016/j.ddtec.2020.11.009',
                'year': 2020
            },
            {
                'num': 9,
                'text': 'Schütt, K. T., Kindermans, P. J., Sauceda, H. E., Chmiela, S., Tkatchenko, A., & Müller, K. R. (2017). SchNet: A continuous-filter convolutional neural network for modeling quantum interactions. NeurIPS 2017.',
                'doi': '10.48550/arXiv.1706.08566',
                'year': 2017
            },
        ]
    }

    # Paper 5: Federated Learning paper
    paper5 = {
        'filename': 'federated_learning_privacy_2023.pdf',
        'title': 'Privacy-Preserving Federated Learning with Adaptive Differential Privacy and Secure Aggregation',
        'authors': 'Ibrahim, A., Chen, L., Johnson, M., & Tanaka, K.',
        'abstract': (
            'Federated learning enables collaborative model training across distributed clients '
            'without sharing raw data. However, model updates can still leak sensitive information. '
            'We propose AdaptDP-Fed, a framework combining adaptive differential privacy with secure '
            'multi-party computation for aggregation. Our method achieves epsilon-delta privacy '
            'guarantees while maintaining 95% of centralized model accuracy.'
        ),
        'references': [
            {
                'num': 1,
                'text': 'McMahan, B., Moore, E., Ramage, D., Hampson, S., & Aguera y Arcas, B. (2017). Communication-efficient learning of deep networks from decentralized data. AISTATS 2017, 1273-1282.',
                'doi': '10.48550/arXiv.1602.05629',
                'year': 2017
            },
            {
                'num': 2,
                'text': 'Dwork, C., McSherry, F., Nissim, K., & Smith, A. (2006). Calibrating noise to sensitivity in private data analysis. TCC 2006, 265-284.',
                'doi': '10.1007/11681878_14',
                'year': 2006
            },
            {
                'num': 3,
                'text': 'Bonawitz, K., Ivanov, V., Kreuter, B., Marcedone, A., McMahan, H. B., Patel, S., ... & Seth, K. (2017). Practical secure aggregation for privacy-preserving machine learning. CCS 2017.',
                'doi': '10.1145/3133956.3133982',
                'year': 2017
            },
            {
                'num': 4,
                'text': 'Geyer, R. C., Klein, T., & Nabi, M. (2017). Differentially private federated learning: A client level perspective. arXiv preprint arXiv:1712.07557.',
                'doi': '10.48550/arXiv.1712.07557',
                'year': 2017
            },
            {
                'num': 5,
                'text': 'Li, T., Sahu, A. K., Talwalkar, A., & Smith, V. (2020). Federated learning: Challenges, methods, and future directions. IEEE Signal Processing Magazine, 37(3), 50-60.',
                'doi': '10.1109/MSP.2020.2975749',
                'year': 2020
            },
            {
                'num': 6,
                'text': 'Kairouz, P., McMahan, H. B., Avent, B., Bellet, A., Bennis, M., Bhagoji, A. N., ... & Zhao, S. (2021). Advances and open problems in federated learning. Foundations and Trends in Machine Learning, 14(1-2), 1-210.',
                'doi': '10.1561/2200000083',
                'year': 2021
            },
            {
                'num': 7,
                'text': 'Wei, K., Li, J., Ding, M., Ma, C., Yang, H. H., Farokhi, F., ... & Poor, H. V. (2020). Federated learning with differential privacy: Algorithms and performance analysis. IEEE Transactions on Information Forensics and Security, 15, 3454-3469.',
                'doi': '10.1109/TIFS.2020.2988575',
                'year': 2020
            },
            {
                'num': 8,
                'text': 'Abadi, M., Chu, A., Goodfellow, I., McMahan, H. B., Mironov, I., Talwar, K., & Zhang, L. (2016). Deep learning with differential privacy. CCS 2016, 308-318.',
                'doi': '10.1145/2976749.2978318',
                'year': 2016
            },
            {
                'num': 9,
                'text': 'Bagdasaryan, E., Veit, A., Hua, Y., Estrin, D., & Shmatikov, V. (2020). How to backdoor federated learning. AISTATS 2020.',
                'doi': '10.48550/arXiv.1807.00459',
                'year': 2020
            },
            {
                'num': 10,
                'text': 'Sun, Z., Kairouz, P., Suresh, A. T., & McMahan, H. B. (2019). Can you really backdoor federated learning? NeurIPS Workshop 2019.',
                'doi': '10.48550/arXiv.1911.07963',
                'year': 2019
            },
        ]
    }

    papers = [paper1, paper2, paper3, paper4, paper5]

    for paper in papers:
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=20)
        pdf.add_page()

        # Title
        pdf.set_font('Helvetica', 'B', 14)
        pdf.set_xy(20, 20)
        pdf.multi_cell(170, 8, paper['title'], align='C')
        pdf.ln(4)

        # Authors
        pdf.set_font('Helvetica', '', 11)
        pdf.multi_cell(170, 6, paper['authors'], align='C')
        pdf.ln(6)

        # Abstract
        pdf.set_font('Helvetica', 'B', 11)
        pdf.cell(0, 6, 'Abstract')
        pdf.ln(4)
        pdf.set_font('Helvetica', '', 10)
        pdf.multi_cell(170, 5, paper['abstract'])
        pdf.ln(8)

        # Introduction placeholder
        pdf.set_font('Helvetica', 'B', 11)
        pdf.cell(0, 6, '1. Introduction')
        pdf.ln(4)
        pdf.set_font('Helvetica', '', 10)
        intro = (
            'Recent advances in deep learning have enabled remarkable progress across '
            'diverse machine learning tasks. This work builds upon established foundations '
            'while introducing novel components that address key limitations of existing approaches. '
            'We demonstrate the effectiveness of our method through extensive experiments on standard benchmarks.'
        )
        pdf.multi_cell(170, 5, intro)
        pdf.ln(6)

        # Method placeholder
        pdf.set_font('Helvetica', 'B', 11)
        pdf.cell(0, 6, '2. Method')
        pdf.ln(4)
        pdf.set_font('Helvetica', '', 10)
        method = (
            'Our approach consists of three main components: (1) the core architecture, '
            '(2) the training objective, and (3) an efficient inference procedure. '
            'We describe each component in detail in the following subsections. '
            'All implementation details and hyperparameter settings are provided in the supplementary material.'
        )
        pdf.multi_cell(170, 5, method)
        pdf.ln(6)

        # Experiments placeholder
        pdf.set_font('Helvetica', 'B', 11)
        pdf.cell(0, 6, '3. Experiments')
        pdf.ln(4)
        pdf.set_font('Helvetica', '', 10)
        exp = (
            'We evaluate our method on several standard benchmarks and compare against '
            'competitive baselines. All experiments are conducted using PyTorch on NVIDIA A100 GPUs. '
            'Results confirm the superiority of our approach in terms of both performance and efficiency.'
        )
        pdf.multi_cell(170, 5, exp)
        pdf.ln(8)

        # References section
        pdf.set_font('Helvetica', 'B', 11)
        pdf.cell(0, 6, 'References')
        pdf.ln(4)
        pdf.set_font('Helvetica', '', 9)

        for ref in paper['references']:
            ref_line = f"[{ref['num']}] {ref['text']}"
            if ref['doi']:
                ref_line += f" DOI: {ref['doi']}"
            pdf.multi_cell(170, 5, ref_line)
            pdf.ln(1)

        output_path = os.path.join(PAPERS_DIR, paper['filename'])
        pdf.output(output_path)
        print(f'Created PDF: {output_path}')

    print(f'All 5 PDFs created in {PAPERS_DIR}')


def create_empty_calc_file():
    """Create an empty LibreOffice Calc file placeholder for the master references."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'References'
    # Just the header row as a starting placeholder
    headers = ['Source_PDF', 'Ref_Number', 'Title', 'Authors', 'Year', 'DOI',
               'DOI_Valid', 'OA_PDF_URL', 'Citation_Count', 'Venue']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    output_path = f'{WORKDIR}/master_references.xlsx'
    wb.save(output_path)
    print(f'Created empty references spreadsheet: {output_path}')
    return output_path


def ensure_documents_dir():
    """Ensure Documents directory exists."""
    docs_dir = f'{WORKDIR}/Documents'
    os.makedirs(docs_dir, exist_ok=True)
    print(f'Ensured Documents directory: {docs_dir}')


def main():
    # Create the papers_to_process folder with PDFs
    create_pdfs()

    # Create a starter spreadsheet
    calc_path = create_empty_calc_file()

    # Ensure Documents directory exists
    ensure_documents_dir()

    # GUI startup: open LibreOffice Calc with the empty references file
    launch_gui(f'libreoffice --calc "{calc_path}"', delay_sec=2.0)
    # Open Chrome for web lookup workflows
    launch_gui('google-chrome --new-window "https://search.crossref.org"', delay_sec=1.5)
    # Open file manager showing papers folder
    launch_gui(f'nautilus "{PAPERS_DIR}"', delay_sec=1.0)

    print('GUI_READY: launched LibreOffice Calc, Chrome, and file manager with DISPLAY=:0')


main()
