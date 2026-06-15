"""
Reward Script: Create comprehensive reference resolution pipeline with master_references.ods,
               bibliography_apa.odt, and bibliography.bib
Task ID: osworld_multi_apps_web_references_015
Domain: libreoffice_calc (multi-app)
Scoring:
  Component 1: master_references.ods — correct columns + data from all 5 PDFs (0.40)
  Component 2: bibliography.bib — BibTeX file with entries in Documents (0.30)
  Component 3: bibliography_apa.odt — APA file with entries sorted alphabetically (0.30)
"""

import os
import re
import shutil
import zipfile

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_web_references_015'

REQUIRED_COLUMNS = [
    'Source_PDF', 'Ref_Number', 'Title', 'Authors', 'Year',
    'DOI', 'DOI_Valid', 'OA_PDF_URL', 'Citation_Count', 'Venue'
]

EXPECTED_SOURCE_PDFS = {
    'attention_transformers_2023.pdf',
    'federated_learning_privacy_2023.pdf',
    'graph_neural_nets_2022.pdf',
    'rl_policy_optimization_2023.pdf',
    'vision_contrastive_2022.pdf',
}

ODS_PATH = os.path.join(WORKDIR, 'master_references.ods')
BIB_PATH = os.path.join(WORKDIR, 'Documents', 'bibliography.bib')
APA_PATH = os.path.join(WORKDIR, 'Documents', 'bibliography_apa.odt')


def load_ods_as_xlsx(ods_path):
    """
    The .ods file may actually be an xlsx-format file with .ods extension.
    Try openpyxl with a renamed copy; fall back gracefully.
    """
    try:
        tmp_path = '/tmp/_reward_ods_check.xlsx'
        shutil.copy(ods_path, tmp_path)
        import openpyxl
        wb = openpyxl.load_workbook(tmp_path)
        return wb, None
    except Exception as e1:
        try:
            # Try real ODF via odfpy
            from odf.opendocument import load as odf_load
            from odf.table import Table, TableRow, TableCell
            from odf.text import P
            return None, odf_load(ods_path)
        except Exception as e2:
            return None, None


def get_para_text(p):
    """Extract text from an ODF paragraph node."""
    text_content = ''
    for node in p.childNodes:
        if hasattr(node, 'data'):
            text_content += node.data
        elif hasattr(node, 'childNodes'):
            for subnode in node.childNodes:
                if hasattr(subnode, 'data'):
                    text_content += subnode.data
    return text_content.strip()


def verify_task():
    total_score = 0.0

    # -------------------------------------------------------------------------
    # Component 1: master_references.ods — data from all 5 PDFs, correct columns
    # (0.40 points)
    # -------------------------------------------------------------------------
    # Sub-component 1a: file exists and has required columns (0.10)
    # Sub-component 1b: has data rows from all 5 source PDFs (0.15)
    # Sub-component 1c: has at least 40 data rows (0.15)
    # -------------------------------------------------------------------------

    if not os.path.exists(ODS_PATH):
        print(f"FAIL: Component 1 — master_references.ods not found at {ODS_PATH}")
    else:
        # Try to load the file
        try:
            wb, odf_doc = load_ods_as_xlsx(ODS_PATH)
            if wb is not None:
                # openpyxl path (xlsx format with .ods extension)
                if 'References' not in wb.sheetnames:
                    print(f"FAIL: Component 1a — 'References' sheet not found, sheets: {wb.sheetnames}")
                else:
                    ws = wb['References']
                    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
                    missing_cols = [col for col in REQUIRED_COLUMNS if col not in headers]
                    if missing_cols:
                        print(f"FAIL: Component 1a — Missing columns: {missing_cols}")
                    else:
                        print(f"PASS: Component 1a — 'References' sheet has all 10 required columns (0.10 pts)")
                        total_score += 0.10

                    # Sub-component 1b: all 5 source PDFs present
                    source_pdfs_in_data = set()
                    data_rows = 0
                    for r in range(2, ws.max_row + 1):
                        src = ws.cell(row=r, column=1).value
                        if src:
                            source_pdfs_in_data.add(src)
                            data_rows += 1

                    missing_pdfs = EXPECTED_SOURCE_PDFS - source_pdfs_in_data
                    if not missing_pdfs:
                        print(f"PASS: Component 1b — Data rows from all 5 PDFs present ({data_rows} total rows) (0.15 pts)")
                        total_score += 0.15
                    else:
                        print(f"FAIL: Component 1b — Missing PDFs in data: {missing_pdfs}. Found: {source_pdfs_in_data}")

                    # Sub-component 1c: at least 40 data rows
                    if data_rows >= 40:
                        print(f"PASS: Component 1c — At least 40 data rows found ({data_rows} rows) (0.15 pts)")
                        total_score += 0.15
                    else:
                        print(f"FAIL: Component 1c — Expected at least 40 data rows, found {data_rows}")

            elif odf_doc is not None:
                # odfpy path (real ODS format)
                from odf.table import Table, TableRow, TableCell
                from odf.text import P
                spreadsheet = odf_doc.spreadsheet
                sheets = spreadsheet.getElementsByType(Table)
                sheet_names = [s.getAttribute('name') for s in sheets]
                ref_sheet = None
                for s in sheets:
                    if s.getAttribute('name') == 'References':
                        ref_sheet = s
                        break
                if ref_sheet is None:
                    print(f"FAIL: Component 1a — 'References' sheet not found. Sheets: {sheet_names}")
                else:
                    rows = ref_sheet.getElementsByType(TableRow)
                    if rows:
                        hrow = rows[0]
                        header_cells = hrow.getElementsByType(TableCell)
                        headers = []
                        for cell in header_cells:
                            ps = cell.getElementsByType(P)
                            val = ''
                            if ps and ps[0].firstChild:
                                val = ps[0].firstChild.data
                            headers.append(val)
                        missing_cols = [col for col in REQUIRED_COLUMNS if col not in headers]
                        if not missing_cols:
                            print(f"PASS: Component 1a — All 10 required columns present (0.10 pts)")
                            total_score += 0.10
                        else:
                            print(f"FAIL: Component 1a — Missing columns: {missing_cols}")

                        source_pdfs_in_data = set()
                        data_rows = 0
                        for row in rows[1:]:
                            cells = row.getElementsByType(TableCell)
                            if cells:
                                ps = cells[0].getElementsByType(P)
                                val = ''
                                if ps and ps[0].firstChild:
                                    val = ps[0].firstChild.data
                                if val:
                                    source_pdfs_in_data.add(val)
                                    data_rows += 1

                        missing_pdfs = EXPECTED_SOURCE_PDFS - source_pdfs_in_data
                        if not missing_pdfs:
                            print(f"PASS: Component 1b — All 5 PDFs represented ({data_rows} rows) (0.15 pts)")
                            total_score += 0.15
                        else:
                            print(f"FAIL: Component 1b — Missing PDFs: {missing_pdfs}")

                        if data_rows >= 40:
                            print(f"PASS: Component 1c — At least 40 data rows ({data_rows} rows) (0.15 pts)")
                            total_score += 0.15
                        else:
                            print(f"FAIL: Component 1c — Only {data_rows} data rows, need 40+")
            else:
                print(f"FAIL: Component 1 — Could not load master_references.ods with openpyxl or odfpy")
        except Exception as e:
            print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: bibliography.bib — BibTeX file with entries in Documents
    # (0.30 points)
    # Sub-component 2a: file exists and has BibTeX structure (0.10)
    # Sub-component 2b: has at least 20 BibTeX entries (0.10)
    # Sub-component 2c: has at least 40 BibTeX entries (0.10)
    # -------------------------------------------------------------------------

    if not os.path.exists(BIB_PATH):
        print(f"FAIL: Component 2 — bibliography.bib not found at {BIB_PATH}")
    else:
        try:
            with open(BIB_PATH, 'r', encoding='utf-8', errors='replace') as f:
                bib_content = f.read()

            # Sub-component 2a: valid BibTeX structure (has @type{ entries)
            bib_entries = re.findall(r'^@\w+\{', bib_content, re.MULTILINE)
            if len(bib_entries) >= 1:
                print(f"PASS: Component 2a — bibliography.bib exists and has BibTeX entries ({len(bib_entries)} total) (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 2a — bibliography.bib has no valid BibTeX entries")

            # Sub-component 2b: at least 20 entries
            if len(bib_entries) >= 20:
                print(f"PASS: Component 2b — At least 20 BibTeX entries found ({len(bib_entries)}) (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 2b — Only {len(bib_entries)} BibTeX entries, need 20+")

            # Sub-component 2c: at least 40 entries
            if len(bib_entries) >= 40:
                print(f"PASS: Component 2c — At least 40 BibTeX entries found ({len(bib_entries)}) (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 2c — Only {len(bib_entries)} BibTeX entries, need 40+")

        except Exception as e:
            print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: bibliography_apa.odt — APA formatted, alphabetically sorted
    # (0.30 points)
    # Sub-component 3a: file exists and contains APA entries (0.10)
    # Sub-component 3b: has at least 20 APA entries (0.10)
    # Sub-component 3c: entries are alphabetically sorted by author (0.10)
    # -------------------------------------------------------------------------

    if not os.path.exists(APA_PATH):
        print(f"FAIL: Component 3 — bibliography_apa.odt not found at {APA_PATH}")
    else:
        try:
            # Try python-docx first (may work if it's actually docx format)
            apa_texts = []
            try:
                from docx import Document as DocxDoc
                docx_tmp = '/tmp/_reward_apa_check.docx'
                shutil.copy(APA_PATH, docx_tmp)
                doc = DocxDoc(docx_tmp)
                apa_texts = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
            except Exception:
                pass

            if not apa_texts:
                # Try odfpy (real ODT format)
                try:
                    from odf.opendocument import load as odf_load
                    from odf.text import P
                    odf_doc = odf_load(APA_PATH)
                    body = odf_doc.text
                    paras = body.getElementsByType(P)
                    apa_texts = [get_para_text(p) for p in paras if get_para_text(p)]
                except Exception as e2:
                    print(f"ERROR: Component 3 — Cannot parse ODT: {e2}")

            if not apa_texts:
                print(f"FAIL: Component 3 — bibliography_apa.odt could not be parsed")
            else:
                # APA entries typically look like: Author, A. (Year). Title. Venue. https://doi.org/...
                # Filter to lines that look like bibliography entries (contain year in parentheses)
                apa_entry_lines = [t for t in apa_texts if re.search(r'\(\d{4}\)', t)]

                # Sub-component 3a: has at least 1 APA entry
                if len(apa_entry_lines) >= 1:
                    print(f"PASS: Component 3a — bibliography_apa.odt has APA entries ({len(apa_entry_lines)} entries) (0.10 pts)")
                    total_score += 0.10
                else:
                    print(f"FAIL: Component 3a — bibliography_apa.odt has no APA-format entries (found {len(apa_texts)} text paragraphs)")

                # Sub-component 3b: at least 20 entries
                if len(apa_entry_lines) >= 20:
                    print(f"PASS: Component 3b — At least 20 APA entries found ({len(apa_entry_lines)}) (0.10 pts)")
                    total_score += 0.10
                else:
                    print(f"FAIL: Component 3b — Only {len(apa_entry_lines)} APA entries, need 20+")

                # Sub-component 3c: alphabetically sorted by first character of each entry
                if len(apa_entry_lines) >= 2:
                    # Extract first author last names for sorting check
                    first_words = []
                    for t in apa_entry_lines:
                        match = re.match(r'^([A-Za-z][^,\s]*)', t)
                        if match:
                            first_words.append(match.group(1).lower())
                    is_sorted = all(first_words[i] <= first_words[i+1] for i in range(len(first_words)-1))
                    if is_sorted:
                        print(f"PASS: Component 3c — APA entries are alphabetically sorted (0.10 pts)")
                        total_score += 0.10
                    else:
                        # Find first out-of-order pair for debugging
                        for i in range(len(first_words)-1):
                            if first_words[i] > first_words[i+1]:
                                print(f"FAIL: Component 3c — Not alphabetically sorted. '{first_words[i]}' before '{first_words[i+1]}'")
                                break
                else:
                    print(f"FAIL: Component 3c — Not enough entries to check alphabetical order ({len(apa_entry_lines)} entries)")

        except Exception as e:
            print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


verify_task()
