"""
Initial Setup: Create a dataset spanning 40+ rows for annual comparison.
Task ID: calc_gsi_019
Domain: libreoffice_calc

The spreadsheet has last year's data in rows 1-20 and this year's data
in rows 21+.  NO split or freeze panes -- the agent must add them.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_019'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Annual Comparison"

    # --- Header row ---
    headers = ["Month", "Region", "Product", "Units Sold", "Revenue ($)",
               "Cost ($)", "Profit ($)", "Growth (%)"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496",
                              fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # --- 2024 Data (rows 2-20, representing last year) ---
    regions = ["North", "South", "East", "West"]
    products = ["Widget A", "Widget B", "Gadget X", "Gadget Y", "Module Z"]
    months_2024 = [
        "Jan 2024", "Feb 2024", "Mar 2024", "Apr 2024", "May 2024",
        "Jun 2024", "Jul 2024", "Aug 2024", "Sep 2024", "Oct 2024",
        "Nov 2024", "Dec 2024", "Jan 2024", "Feb 2024", "Mar 2024",
        "Apr 2024", "May 2024", "Jun 2024", "Jul 2024",
    ]

    data_2024 = [
        ["Jan 2024", "North", "Widget A", 152, 45230.00, 28140.00, 17090.00, 3.2],
        ["Feb 2024", "South", "Gadget X", 98, 32670.50, 21450.00, 11220.50, -1.5],
        ["Mar 2024", "East", "Widget B", 214, 58920.00, 35600.00, 23320.00, 5.8],
        ["Apr 2024", "West", "Module Z", 67, 24510.75, 16200.00, 8310.75, 2.1],
        ["May 2024", "North", "Gadget Y", 183, 51840.00, 33100.00, 18740.00, 4.6],
        ["Jun 2024", "South", "Widget A", 129, 38470.25, 24800.00, 13670.25, -0.3],
        ["Jul 2024", "East", "Gadget X", 241, 67350.00, 42100.00, 25250.00, 7.2],
        ["Aug 2024", "West", "Widget B", 88, 29180.50, 18900.00, 10280.50, 1.8],
        ["Sep 2024", "North", "Module Z", 176, 49720.00, 31400.00, 18320.00, 3.9],
        ["Oct 2024", "South", "Gadget Y", 113, 35460.75, 22700.00, 12760.75, -2.1],
        ["Nov 2024", "East", "Widget A", 198, 54890.00, 34200.00, 20690.00, 6.4],
        ["Dec 2024", "West", "Gadget X", 72, 21930.25, 14500.00, 7430.25, 0.7],
        ["Jan 2024", "North", "Widget B", 145, 42310.00, 27800.00, 14510.00, 2.5],
        ["Feb 2024", "South", "Module Z", 91, 28750.50, 19100.00, 9650.50, -1.2],
        ["Mar 2024", "East", "Gadget Y", 223, 61540.00, 38400.00, 23140.00, 5.1],
        ["Apr 2024", "West", "Widget A", 56, 19870.75, 13200.00, 6670.75, -3.4],
        ["May 2024", "North", "Gadget X", 167, 47290.00, 30100.00, 17190.00, 4.0],
        ["Jun 2024", "South", "Widget B", 104, 33120.25, 21600.00, 11520.25, 1.3],
        ["Jul 2024", "East", "Module Z", 189, 52680.00, 33800.00, 18880.00, 6.9],
    ]

    # --- Section divider row for 2024 ---
    # Row 2 starts 2024 data through row 20 (19 data rows)

    for r, row_data in enumerate(data_2024, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = header_border
            if c in (5, 6, 7):  # currency columns
                cell.number_format = '$#,##0.00'
            elif c == 8:  # percentage
                cell.number_format = '0.0'

    # --- 2025 Data (rows 21-40, representing this year) ---
    data_2025 = [
        ["Jan 2025", "North", "Widget A", 168, 49850.00, 30200.00, 19650.00, 4.1],
        ["Feb 2025", "South", "Gadget X", 107, 35420.50, 23100.00, 12320.50, 0.8],
        ["Mar 2025", "East", "Widget B", 232, 63710.00, 38400.00, 25310.00, 6.5],
        ["Apr 2025", "West", "Module Z", 74, 26830.75, 17500.00, 9330.75, 3.2],
        ["May 2025", "North", "Gadget Y", 195, 55120.00, 35200.00, 19920.00, 5.3],
        ["Jun 2025", "South", "Widget A", 138, 41230.25, 26500.00, 14730.25, 1.2],
        ["Jul 2025", "East", "Gadget X", 259, 72410.00, 45200.00, 27210.00, 8.1],
        ["Aug 2025", "West", "Widget B", 96, 31540.50, 20400.00, 11140.50, 2.7],
        ["Sep 2025", "North", "Module Z", 188, 52910.00, 33600.00, 19310.00, 4.8],
        ["Oct 2025", "South", "Gadget Y", 124, 38790.75, 24800.00, 13990.75, -0.5],
        ["Nov 2025", "East", "Widget A", 215, 59340.00, 36800.00, 22540.00, 7.3],
        ["Dec 2025", "West", "Gadget X", 81, 24670.25, 16100.00, 8570.25, 1.9],
        ["Jan 2025", "North", "Widget B", 159, 46120.00, 29800.00, 16320.00, 3.6],
        ["Feb 2025", "South", "Module Z", 102, 31480.50, 20700.00, 10780.50, 0.4],
        ["Mar 2025", "East", "Gadget Y", 241, 66280.00, 41200.00, 25080.00, 6.1],
        ["Apr 2025", "West", "Widget A", 63, 22140.75, 14600.00, 7540.75, -2.1],
        ["May 2025", "North", "Gadget X", 179, 50530.00, 32400.00, 18130.00, 4.9],
        ["Jun 2025", "South", "Widget B", 115, 36250.25, 23400.00, 12850.25, 2.2],
        ["Jul 2025", "East", "Module Z", 204, 56890.00, 36100.00, 20790.00, 7.6],
        ["Aug 2025", "West", "Gadget Y", 87, 27310.50, 17800.00, 9510.50, 1.1],
    ]

    for r, row_data in enumerate(data_2025, 21):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = header_border
            if c in (5, 6, 7):
                cell.number_format = '$#,##0.00'
            elif c == 8:
                cell.number_format = '0.0'

    # --- Column widths ---
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 13
    ws.column_dimensions["G"].width = 13
    ws.column_dimensions["H"].width = 12

    # NO split or freeze panes -- agent must do this
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
