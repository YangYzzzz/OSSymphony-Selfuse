"""
Reward Script: Employee Satisfaction Survey Analysis Dashboard
Task ID: calc_hr_053
Domain: libreoffice_calc

Scoring Rubric:
  Component 1 (0.35): AVERAGEIF formulas in B4:B7 for satisfaction by department
  Component 2 (0.35): AVERAGEIFS formulas in B12:B15 for engagement (tenure >3) by dept
  Component 3 (0.30): COUNTIFS formulas in B20:B23 for WLB < 3 count by dept

All components verify formulas that do NOT exist in the initial state (cells are empty).
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_053'
FILE_PATH = os.path.join(WORKDIR, f'{TASK_ID}.xlsx')

# Persistence hook: save any unsaved LibreOffice state
def persist_app_state():
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        import time
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def has_formula(value, func_name):
    """Check if a cell value is a formula containing the given function name."""
    if not isinstance(value, str):
        return False
    return value.strip().startswith('=') and func_name.upper() in value.upper()


def check_averageif_not_averageifs(value):
    """Check that a formula uses AVERAGEIF but not AVERAGEIFS (single-criteria)."""
    if not isinstance(value, str):
        return False
    upper = value.upper().replace(" ", "")
    # Must contain AVERAGEIF and start with =
    if not upper.startswith('=') or 'AVERAGEIF' not in upper:
        return False
    # Accept: contains AVERAGEIF but not AVERAGEIFS
    # Also accept AVERAGEIFS since it's a superset that also works
    return True


DEPARTMENTS = ['Engineering', 'Sales', 'HR', 'Marketing']


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify Dashboard sheet exists
    if 'Dashboard' not in wb.sheetnames:
        print("FAIL: 'Dashboard' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Dashboard']

    # ================================================================
    # Component 1: AVERAGEIF formulas for avg satisfaction by department
    # Cells B4:B7 should contain AVERAGEIF formulas referencing Survey data
    # (0.35 points — 4 cells, each worth ~0.0875)
    # ================================================================
    try:
        comp1_score = 0.0
        comp1_cells = {'B4': 'Engineering', 'B5': 'Sales', 'B6': 'HR', 'B7': 'Marketing'}
        comp1_pass_count = 0

        for cell_ref, dept in comp1_cells.items():
            val = ws[cell_ref].value
            if val is not None and isinstance(val, str) and val.strip().startswith('='):
                upper_val = val.upper().replace(" ", "")
                # Must use AVERAGEIF (or AVERAGEIFS) and reference the department
                if 'AVERAGEIF' in upper_val:
                    # Check it references satisfaction column (D) and department criterion
                    if dept.upper() in upper_val or f'"{dept.upper()}"' in upper_val or dept.upper() in val.upper():
                        comp1_pass_count += 1
                        print(f"  PASS: {cell_ref} has AVERAGEIF formula for {dept}: {val}")
                    else:
                        print(f"  PARTIAL: {cell_ref} has AVERAGEIF but missing dept reference for {dept}: {val}")
                        comp1_pass_count += 0.5
                else:
                    print(f"  FAIL: {cell_ref} has formula but not AVERAGEIF: {val}")
            else:
                print(f"  FAIL: {cell_ref} is empty or not a formula: {val}")

        comp1_score = 0.35 * (comp1_pass_count / 4)
        if comp1_pass_count > 0:
            print(f"PASS: Component 1 — AVERAGEIF formulas ({comp1_pass_count}/4 cells) ({comp1_score:.3f} pts)")
            total_score += comp1_score
        else:
            print(f"FAIL: Component 1 — No AVERAGEIF formulas found in B4:B7")

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ================================================================
    # Component 2: AVERAGEIFS formulas for avg engagement (tenure >3) by dept
    # Cells B12:B15 should contain AVERAGEIFS formulas with multi-criteria
    # (0.35 points — 4 cells, each worth ~0.0875)
    # ================================================================
    try:
        comp2_score = 0.0
        comp2_cells = {'B12': 'Engineering', 'B13': 'Sales', 'B14': 'HR', 'B15': 'Marketing'}
        comp2_pass_count = 0

        for cell_ref, dept in comp2_cells.items():
            val = ws[cell_ref].value
            if val is not None and isinstance(val, str) and val.strip().startswith('='):
                upper_val = val.upper().replace(" ", "")
                # Must use AVERAGEIFS (multi-criteria) and reference tenure >3
                if 'AVERAGEIFS' in upper_val:
                    # Check it has tenure criterion (>3) and department
                    has_tenure_crit = ('>3' in upper_val or '>"&3' in upper_val or
                                       '">"&3' in upper_val or '>3' in val)
                    if has_tenure_crit:
                        comp2_pass_count += 1
                        print(f"  PASS: {cell_ref} has AVERAGEIFS formula with tenure >3 for {dept}: {val}")
                    else:
                        print(f"  PARTIAL: {cell_ref} has AVERAGEIFS but missing tenure >3 criterion: {val}")
                        comp2_pass_count += 0.5
                elif 'AVERAGEIF' in upper_val:
                    # Using single-criteria AVERAGEIF when AVERAGEIFS is needed
                    print(f"  PARTIAL: {cell_ref} uses AVERAGEIF instead of AVERAGEIFS: {val}")
                    comp2_pass_count += 0.25
                else:
                    print(f"  FAIL: {cell_ref} has formula but not AVERAGEIFS: {val}")
            else:
                print(f"  FAIL: {cell_ref} is empty or not a formula: {val}")

        comp2_score = 0.35 * (comp2_pass_count / 4)
        if comp2_pass_count > 0:
            print(f"PASS: Component 2 — AVERAGEIFS formulas ({comp2_pass_count}/4 cells) ({comp2_score:.3f} pts)")
            total_score += comp2_score
        else:
            print(f"FAIL: Component 2 — No AVERAGEIFS formulas found in B12:B15")

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ================================================================
    # Component 3: COUNTIFS formulas for WLB < 3 count by department
    # Cells B20:B23 should contain COUNTIFS formulas
    # (0.30 points — 4 cells, each worth 0.075)
    # ================================================================
    try:
        comp3_score = 0.0
        comp3_cells = {'B20': 'Engineering', 'B21': 'Sales', 'B22': 'HR', 'B23': 'Marketing'}
        comp3_pass_count = 0

        for cell_ref, dept in comp3_cells.items():
            val = ws[cell_ref].value
            if val is not None and isinstance(val, str) and val.strip().startswith('='):
                upper_val = val.upper().replace(" ", "")
                # Must use COUNTIFS and reference WLB < 3
                if 'COUNTIFS' in upper_val:
                    # Check it has <3 criterion for WLB
                    has_wlb_crit = ('<3' in upper_val or '<"&3' in upper_val or
                                    '"<"&3' in upper_val or '<3' in val)
                    if has_wlb_crit:
                        comp3_pass_count += 1
                        print(f"  PASS: {cell_ref} has COUNTIFS formula with WLB<3 for {dept}: {val}")
                    else:
                        print(f"  PARTIAL: {cell_ref} has COUNTIFS but missing WLB<3 criterion: {val}")
                        comp3_pass_count += 0.5
                elif 'COUNTIF' in upper_val:
                    # Using single-criteria COUNTIF when COUNTIFS is needed
                    print(f"  PARTIAL: {cell_ref} uses COUNTIF instead of COUNTIFS: {val}")
                    comp3_pass_count += 0.25
                else:
                    print(f"  FAIL: {cell_ref} has formula but not COUNTIFS: {val}")
            else:
                print(f"  FAIL: {cell_ref} is empty or not a formula: {val}")

        comp3_score = 0.30 * (comp3_pass_count / 4)
        if comp3_pass_count > 0:
            print(f"PASS: Component 3 — COUNTIFS formulas ({comp3_pass_count}/4 cells) ({comp3_score:.3f} pts)")
            total_score += comp3_score
        else:
            print(f"FAIL: Component 3 — No COUNTIFS formulas found in B20:B23")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.3f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state()

if not os.path.exists(FILE_PATH):
    print(f"File not found: {FILE_PATH}")
    print("REWARD: 0.0")
else:
    verify_task(FILE_PATH)
