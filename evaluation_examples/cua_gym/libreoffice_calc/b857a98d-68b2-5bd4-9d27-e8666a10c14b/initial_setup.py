"""
Initial Setup: Employee satisfaction survey analysis workbook
Task ID: calc_hr_053
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_053'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Survey ---
    ws = wb.active
    ws.title = 'Survey'

    headers = ['Employee', 'Department', 'Tenure', 'Satisfaction', 'Engagement', 'WLB Score']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.font = Font(bold=True, size=11, color="FFFFFF")

    # 30 rows of realistic employee survey data
    data = [
        ['Sarah Chen', 'Engineering', 5, 4.2, 4.5, 3.8],
        ['Marcus Johnson', 'Sales', 2, 3.5, 3.2, 2.8],
        ['Priya Patel', 'Engineering', 7, 4.8, 4.7, 4.1],
        ['David Kim', 'HR', 3, 3.9, 3.6, 3.5],
        ['Emily Rodriguez', 'Marketing', 1, 4.1, 4.0, 3.2],
        ['James Wilson', 'Engineering', 4, 3.7, 4.1, 2.5],
        ['Aisha Mohammed', 'Sales', 6, 3.3, 3.0, 2.1],
        ['Robert Taylor', 'HR', 2, 4.5, 4.2, 4.0],
        ['Lisa Wang', 'Marketing', 8, 4.0, 3.8, 3.6],
        ['Michael Brown', 'Engineering', 3, 4.4, 4.6, 3.9],
        ['Jennifer Lee', 'Sales', 5, 3.8, 3.5, 2.9],
        ['Carlos Garcia', 'HR', 1, 3.2, 3.1, 2.7],
        ['Amanda Foster', 'Engineering', 6, 4.6, 4.8, 4.2],
        ['Daniel Thompson', 'Marketing', 4, 3.6, 3.4, 3.0],
        ['Rachel Green', 'Sales', 3, 4.0, 3.7, 3.3],
        ['Kevin O\'Brien', 'Engineering', 2, 3.9, 4.0, 2.6],
        ['Sophia Martinez', 'HR', 5, 4.3, 4.1, 3.7],
        ['Tyler Jackson', 'Marketing', 7, 3.4, 3.2, 2.4],
        ['Nicole Adams', 'Sales', 4, 3.1, 2.9, 2.2],
        ['Christopher Lee', 'Engineering', 8, 4.7, 4.9, 4.3],
        ['Maria Gonzalez', 'HR', 3, 3.8, 3.5, 3.1],
        ['Andrew White', 'Marketing', 2, 4.2, 4.0, 3.4],
        ['Jessica Huang', 'Sales', 6, 3.6, 3.3, 2.5],
        ['Patrick Murphy', 'Engineering', 1, 4.0, 4.3, 3.0],
        ['Lauren Davis', 'HR', 4, 4.1, 3.9, 3.8],
        ['Ryan Nakamura', 'Marketing', 5, 3.5, 3.6, 2.8],
        ['Olivia Scott', 'Sales', 2, 3.7, 3.4, 3.1],
        ['Brandon Reed', 'Engineering', 9, 4.5, 4.4, 4.0],
        ['Hannah Miller', 'HR', 6, 3.6, 3.3, 2.9],
        ['Victor Perez', 'Marketing', 3, 4.3, 4.2, 3.5],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12

    # --- Sheet 2: Dashboard (empty, ready for formulas) ---
    ws2 = wb.create_sheet('Dashboard')

    # Section 1: Average Satisfaction by Department
    ws2['A1'] = 'Average Satisfaction by Department'
    ws2['A1'].font = Font(bold=True, size=13)

    ws2['A3'] = 'Department'
    ws2['B3'] = 'Avg Satisfaction'
    ws2['A3'].font = Font(bold=True)
    ws2['B3'].font = Font(bold=True)

    ws2['A4'] = 'Engineering'
    ws2['A5'] = 'Sales'
    ws2['A6'] = 'HR'
    ws2['A7'] = 'Marketing'

    # B4:B7 left EMPTY - task is to add AVERAGEIF formulas here

    # Section 2: Avg Engagement for Tenure >3 years by Department
    ws2['A9'] = 'Avg Engagement (Tenure >3 years) by Department'
    ws2['A9'].font = Font(bold=True, size=13)

    ws2['A11'] = 'Department'
    ws2['B11'] = 'Avg Engagement (Tenure >3)'
    ws2['A11'].font = Font(bold=True)
    ws2['B11'].font = Font(bold=True)

    ws2['A12'] = 'Engineering'
    ws2['A13'] = 'Sales'
    ws2['A14'] = 'HR'
    ws2['A15'] = 'Marketing'

    # B12:B15 left EMPTY - task is to add AVERAGEIFS formulas here

    # Section 3: Count of employees with WLB Score < 3 by Department
    ws2['A17'] = 'Employees with WLB Score Below 3 by Department'
    ws2['A17'].font = Font(bold=True, size=13)

    ws2['A19'] = 'Department'
    ws2['B19'] = 'Count (WLB < 3)'
    ws2['A19'].font = Font(bold=True)
    ws2['B19'].font = Font(bold=True)

    ws2['A20'] = 'Engineering'
    ws2['A21'] = 'Sales'
    ws2['A22'] = 'HR'
    ws2['A23'] = 'Marketing'

    # B20:B23 left EMPTY - task is to add COUNTIFS formulas here

    # Set column widths
    ws2.column_dimensions['A'].width = 42
    ws2.column_dimensions['B'].width = 28

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
