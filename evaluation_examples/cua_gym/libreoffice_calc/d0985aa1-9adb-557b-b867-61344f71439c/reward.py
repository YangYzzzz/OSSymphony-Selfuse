"""
Reward Script: Add percentage data labels to pie chart and move legend to bottom
Task ID: calc_gg2_020
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Data labels show percentage (showPercent=True)
  Component 2 (0.35): Data labels show category name with separator (showCatName=True + separator)
  Component 3 (0.35): Legend position changed to bottom ('b')
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg2_020'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice documents."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def get_chart_data_labels(chart):
    """
    Get data labels from chart-level or series-level.
    Returns the DataLabelList object or None.
    """
    # Check chart-level dataLabels first
    if hasattr(chart, 'dataLabels') and chart.dataLabels is not None:
        return chart.dataLabels
    # Check series-level dLbls
    if chart.series and len(chart.series) > 0:
        s = chart.series[0]
        if hasattr(s, 'dLbls') and s.dLbls is not None:
            return s.dLbls
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Market Share' sheet must exist
    if 'Market Share' not in wb.sheetnames:
        print(f"CRITICAL: 'Market Share' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Market Share']

    # Precondition: must have at least one chart
    charts = ws._charts
    if not charts or len(charts) == 0:
        print("CRITICAL: No charts found on 'Market Share' sheet")
        print("REWARD: 0.0")
        return 0.0

    chart = charts[0]

    # Component 1: Data labels show percentage (0.30 points)
    # Initial state: no data labels at all → this FAILS on initial
    try:
        dl = get_chart_data_labels(chart)
        if dl is not None and dl.showPercent is True:
            print(f"PASS: Component 1 — Data labels showPercent=True (0.30 pts)")
            total_score += 0.30
        else:
            show_pct = dl.showPercent if dl is not None else None
            print(f"FAIL: Component 1 — Expected showPercent=True, found: {show_pct} (dl={dl is not None})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Data labels show category name with separator (0.35 points)
    # Initial state: no data labels → this FAILS on initial
    try:
        dl = get_chart_data_labels(chart)
        if dl is not None and dl.showCatName is True:
            # Check for separator (': ' or similar)
            sep = dl.separator if hasattr(dl, 'separator') else None
            if sep is not None and len(str(sep).strip()) > 0:
                print(f"PASS: Component 2 — Data labels showCatName=True with separator='{sep}' (0.35 pts)")
                total_score += 0.35
            else:
                # Category name shown but no explicit separator — partial credit
                if dl.showCatName is True:
                    print(f"PARTIAL: Component 2 — showCatName=True but no separator (separator={sep}). (0.20 pts)")
                    total_score += 0.20
        else:
            show_cat = dl.showCatName if dl is not None else None
            print(f"FAIL: Component 2 — Expected showCatName=True, found: {show_cat}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Legend position is bottom (0.35 points)
    # Initial state: legend position is 'r' (right) → this FAILS on initial
    try:
        legend = chart.legend
        if legend is not None:
            pos = legend.position
            if pos == 'b':
                print(f"PASS: Component 3 — Legend position='b' (bottom) (0.35 pts)")
                total_score += 0.35
            else:
                print(f"FAIL: Component 3 — Expected legend position='b', found: '{pos}'")
        else:
            print(f"FAIL: Component 3 — No legend found on chart")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
