"""
Initial Setup: Pie chart with no data labels and legend on right
Task ID: calc_gg2_020
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.legend import Legend

WORKDIR = '/home/user'
TASK_ID = 'calc_gg2_020'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Market Share sheet ---
    ws = wb.active
    ws.title = 'Market Share'

    # Headers
    ws.cell(row=1, column=1, value='Region')
    ws.cell(row=1, column=2, value='Market Share (%)')

    # Data: five regions
    data = [
        ['North', 34],
        ['South', 22],
        ['East', 18],
        ['West', 15],
        ['International', 11],
    ]
    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        ws.cell(row=r, column=2, value=row_data[1])

    # Column widths for readability
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18

    # Create pie chart with NO data labels, legend on RIGHT (default)
    pie = PieChart()
    pie.title = 'Regional Market Share 2024'
    pie.style = 10

    cat_data = Reference(ws, min_col=1, min_row=2, max_row=6)
    val_data = Reference(ws, min_col=2, min_row=1, max_row=6)
    pie.add_data(val_data, titles_from_data=True)
    pie.set_categories(cat_data)

    # Explicitly set legend to the right side (default but being explicit)
    pie.legend = Legend()
    pie.legend.position = 'r'  # right side

    # No data labels (default - we do NOT set any dataLabels)

    # Size the chart
    pie.width = 18
    pie.height = 14

    ws.add_chart(pie, 'D2')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
