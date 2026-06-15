"""
Initial Setup: Create a spreadsheet with project data and deadline dates.
Task ID: calc_nrv_040
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
from datetime import date

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_040'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Headers
    ws["A1"] = "Project"
    ws["B1"] = "Manager"
    ws["C1"] = "Deadline"

    # E1 label; E2 and E3 intentionally empty
    ws["E1"] = "Date Stats"

    # Project data rows 2-20 (19 rows)
    projects = [
        ("Website Redesign", "Sarah Chen", date(2024, 3, 15)),
        ("Mobile App v2", "Marcus Johnson", date(2024, 4, 1)),
        ("Data Migration", "Priya Sharma", date(2024, 4, 22)),
        ("Cloud Infrastructure", "David Park", date(2024, 5, 10)),
        ("API Gateway", "Elena Rodriguez", date(2024, 5, 30)),
        ("Security Audit", "James O'Brien", date(2024, 6, 15)),
        ("Customer Portal", "Aisha Patel", date(2024, 7, 1)),
        ("Analytics Dashboard", "Robert Kim", date(2024, 7, 20)),
        ("Payment Integration", "Lisa Nakamura", date(2024, 8, 5)),
        ("Inventory System", "Michael Torres", date(2024, 8, 28)),
        ("HR Onboarding Tool", "Fatima Al-Hassan", date(2024, 9, 12)),
        ("Supply Chain Tracker", "Chris Anderson", date(2024, 9, 30)),
        ("Email Campaign Engine", "Nina Petrova", date(2024, 10, 15)),
        ("Compliance Dashboard", "Thomas Wright", date(2024, 10, 31)),
        ("CRM Enhancement", "Yuki Tanaka", date(2024, 11, 8)),
        ("Warehouse Automation", "Patrick Murphy", date(2024, 11, 22)),
        ("Fraud Detection ML", "Ananya Gupta", date(2024, 12, 5)),
        ("Vendor Management", "Daniel Cohen", date(2024, 12, 18)),
        ("Year-End Reporting", "Grace Liu", date(2024, 12, 31)),
    ]

    for r, (project, manager, deadline) in enumerate(projects, 2):
        ws.cell(row=r, column=1, value=project)
        ws.cell(row=r, column=2, value=manager)
        ws.cell(row=r, column=3, value=deadline)

    # Format date column
    for r in range(2, 21):
        ws.cell(row=r, column=3).number_format = 'yyyy-mm-dd'

    # Adjust column widths for readability
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["E"].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
