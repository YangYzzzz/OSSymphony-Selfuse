"""
Reward Script: Create named range 'ProjectDates' and MIN/MAX formulas
Task ID: calc_nrv_040
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Named range 'ProjectDates' exists and refers to Sheet1!$C$2:$C$20
  Component 2 (0.3): E2 contains =MIN(ProjectDates) formula
  Component 3 (0.3): E3 contains =MAX(ProjectDates) formula
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_040'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Component 1: Named range 'ProjectDates' exists and refers to $C$2:$C$20 (0.4 points)
    try:
        defined_names = dict(wb.defined_names)
        if 'ProjectDates' in defined_names:
            dn = defined_names['ProjectDates']
            ref_value = dn.attr_text
            # Normalize: accept variations like Sheet1!$C$2:$C$20
            normalized = ref_value.upper().replace("'", "").replace(" ", "")
            if "$C$2:$C$20" in normalized:
                print(f"PASS: Component 1 — Named range 'ProjectDates' = {ref_value} (0.4 pts)")
                total_score += 0.4
            else:
                print(f"FAIL: Component 1 — Named range 'ProjectDates' exists but refers to {ref_value}, expected $C$2:$C$20")
        else:
            print(f"FAIL: Component 1 — Named range 'ProjectDates' not found. Defined names: {list(defined_names.keys())}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: E2 contains =MIN(ProjectDates) formula (0.3 points)
    try:
        e2_value = ws['E2'].value
        if isinstance(e2_value, str):
            e2_normalized = e2_value.upper().replace(" ", "")
            if e2_normalized == "=MIN(PROJECTDATES)":
                print(f"PASS: Component 2 — E2 contains {e2_value} (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — E2 contains '{e2_value}', expected =MIN(ProjectDates)")
        else:
            print(f"FAIL: Component 2 — E2 value is {repr(e2_value)} (type: {type(e2_value).__name__}), expected formula =MIN(ProjectDates)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: E3 contains =MAX(ProjectDates) formula (0.3 points)
    try:
        e3_value = ws['E3'].value
        if isinstance(e3_value, str):
            e3_normalized = e3_value.upper().replace(" ", "")
            if e3_normalized == "=MAX(PROJECTDATES)":
                print(f"PASS: Component 3 — E3 contains {e3_value} (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 3 — E3 contains '{e3_value}', expected =MAX(ProjectDates)")
        else:
            print(f"FAIL: Component 3 — E3 value is {repr(e3_value)} (type: {type(e3_value).__name__}), expected formula =MAX(ProjectDates)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 1)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
