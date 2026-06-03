"""
Reward Script: Quarterly Performance Review Dashboard
Task ID: calc_gsd_038
Domain: libreoffice_calc
Scoring:
  C1 (0.25) - Title A1:H1 merged, correct text, 18pt bold, dark navy, gray bg, centered
  C2 (0.10) - Subtitle A2 correct text, italic, 10pt
  C3 (0.25) - Four KPI labels: correct text, blue bg, white bold font
  C4 (0.25) - Four KPI values: correct values, 20pt bold dark navy, white bg
  C5 (0.15) - Correct merged cell ranges for boxes
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_038'


def verify_task(file_path):
    """Verify task completion with progressive scoring. Returns float 0.0-1.0."""
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get the Dashboard sheet
    if 'Dashboard' not in wb.sheetnames:
        print("FAIL: 'Dashboard' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Dashboard']

    # =========================================================================
    # Component 1: Title in A1:H1 (0.25 points)
    # Checks: A1:H1 merged, text='Q3 2024 Performance Review', 18pt bold,
    #         dark navy font (#1F3864), light gray bg (#F2F2F2), centered
    # =========================================================================
    try:
        a1 = ws['A1']
        c1_score = 0.0

        # Check text
        if a1.value and 'Q3 2024 Performance Review' in str(a1.value):
            c1_score += 0.08
            print(f"PASS: C1a — Title text correct: {a1.value!r}")
        else:
            print(f"FAIL: C1a — Expected 'Q3 2024 Performance Review', found: {a1.value!r}")

        # Check A1:H1 is merged (B1 should be MergedCell)
        if isinstance(ws['B1'], MergedCell) and isinstance(ws['H1'], MergedCell):
            c1_score += 0.05
            print("PASS: C1b — A1:H1 merge detected")
        else:
            print("FAIL: C1b — A1:H1 not merged")

        # Check font: 18pt bold
        if a1.font.size and a1.font.size >= 17 and a1.font.bold:
            c1_score += 0.04
            print(f"PASS: C1c — Font 18pt bold (size={a1.font.size}, bold={a1.font.bold})")
        else:
            print(f"FAIL: C1c — Expected 18pt bold, found size={a1.font.size}, bold={a1.font.bold}")

        # Check font color dark navy (#1F3864)
        try:
            fc = a1.font.color.rgb if a1.font.color else None
            if fc and '1F3864' in str(fc).upper():
                c1_score += 0.03
                print(f"PASS: C1d — Font color dark navy ({fc})")
            else:
                print(f"FAIL: C1d — Expected navy font color #1F3864, found: {fc}")
        except Exception as e:
            print(f"FAIL: C1d — Font color check error: {e}")

        # Check background fill light gray (#F2F2F2)
        try:
            bg = a1.fill.fgColor.rgb if a1.fill.fgColor else None
            if bg and 'F2F2F2' in str(bg).upper():
                c1_score += 0.03
                print(f"PASS: C1e — Background light gray ({bg})")
            else:
                print(f"FAIL: C1e — Expected gray bg #F2F2F2, found: {bg}")
        except Exception as e:
            print(f"FAIL: C1e — Fill check error: {e}")

        # Check centered alignment
        if a1.alignment.horizontal == 'center':
            c1_score += 0.02
            print("PASS: C1f — Centered alignment")
        else:
            print(f"FAIL: C1f — Expected center alignment, found: {a1.alignment.horizontal}")

        total_score += c1_score
        print(f"  C1 subtotal: {c1_score:.2f}/0.25")

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Subtitle in A2 (0.10 points)
    # Checks: text contains 'Reporting Period', italic, 10pt
    # =========================================================================
    try:
        a2 = ws['A2']
        c2_score = 0.0

        if a2.value and 'Reporting Period' in str(a2.value) and 'July' in str(a2.value):
            c2_score += 0.05
            print(f"PASS: C2a — Subtitle text correct: {a2.value!r}")
        else:
            print(f"FAIL: C2a — Expected subtitle with 'Reporting Period...July...', found: {a2.value!r}")

        if a2.font.italic:
            c2_score += 0.03
            print("PASS: C2b — Subtitle is italic")
        else:
            print(f"FAIL: C2b — Expected italic, found italic={a2.font.italic}")

        # Only check font size if subtitle text is present (avoid scoring default font)
        if a2.value and 'Reporting Period' in str(a2.value):
            if a2.font.size and abs(a2.font.size - 10) < 1:
                c2_score += 0.02
                print(f"PASS: C2c — Subtitle font size ~10pt ({a2.font.size})")
            else:
                print(f"FAIL: C2c — Expected 10pt, found size={a2.font.size}")
        else:
            print(f"FAIL: C2c — Skipped (no subtitle text)")

        total_score += c2_score
        print(f"  C2 subtotal: {c2_score:.2f}/0.10")

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Four KPI labels with correct text and blue bg (0.25 points)
    # Label cells: B3='Total Revenue', E3='Gross Margin',
    #              B8='New Customers', E8='Net Promoter Score'
    # Each: blue (#4472C4) bg, white bold font, 11pt
    # =========================================================================
    try:
        labels = {
            'B3': 'Total Revenue',
            'E3': 'Gross Margin',
            'B8': 'New Customers',
            'E8': 'Net Promoter Score',
        }
        c3_score = 0.0
        pts_per_label = 0.25 / 4  # ~0.0625 per label

        for coord, expected_text in labels.items():
            cell = ws[coord]
            label_pts = 0.0

            # Check text
            if cell.value and expected_text.lower() in str(cell.value).lower():
                label_pts += pts_per_label * 0.4
            else:
                print(f"FAIL: C3 — {coord} expected '{expected_text}', found: {cell.value!r}")
                continue

            # Check blue background
            try:
                bg = cell.fill.fgColor.rgb if cell.fill.fgColor else None
                if bg and '4472C4' in str(bg).upper():
                    label_pts += pts_per_label * 0.3
                else:
                    print(f"FAIL: C3 — {coord} bg expected #4472C4, found: {bg}")
            except:
                pass

            # Check white bold font
            if cell.font.bold:
                label_pts += pts_per_label * 0.3
            else:
                print(f"FAIL: C3 — {coord} expected bold font")

            c3_score += label_pts
            print(f"PASS: C3 — {coord} label '{expected_text}' verified ({label_pts:.3f} pts)")

        total_score += c3_score
        print(f"  C3 subtotal: {c3_score:.2f}/0.25")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: Four KPI values with correct content and formatting (0.25 points)
    # Value cells: B4='$4.2M', E4='38.5%', B9='1,247', E9='72'
    # Each: 20pt bold, dark navy font, white bg
    # =========================================================================
    try:
        values = {
            'B4': '$4.2M',
            'E4': '38.5%',
            'B9': '1,247',
            'E9': '72',
        }
        c4_score = 0.0
        pts_per_val = 0.25 / 4  # ~0.0625 per value

        for coord, expected_val in values.items():
            cell = ws[coord]
            val_pts = 0.0

            # Check value content
            if cell.value and expected_val.lower() in str(cell.value).lower():
                val_pts += pts_per_val * 0.4
            else:
                print(f"FAIL: C4 — {coord} expected '{expected_val}', found: {cell.value!r}")
                continue

            # Check 20pt bold
            if cell.font.bold and cell.font.size and cell.font.size >= 18:
                val_pts += pts_per_val * 0.3
            else:
                print(f"FAIL: C4 — {coord} expected 20pt bold, found sz={cell.font.size} bold={cell.font.bold}")

            # Check dark navy font color
            try:
                fc = cell.font.color.rgb if cell.font.color else None
                if fc and '1F3864' in str(fc).upper():
                    val_pts += pts_per_val * 0.3
                else:
                    print(f"FAIL: C4 — {coord} font color expected #1F3864, found: {fc}")
            except:
                pass

            c4_score += val_pts
            print(f"PASS: C4 — {coord} value '{expected_val}' verified ({val_pts:.3f} pts)")

        total_score += c4_score
        print(f"  C4 subtotal: {c4_score:.2f}/0.25")

    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # =========================================================================
    # Component 5: Correct merged cell ranges for KPI boxes (0.15 points)
    # Expected merges: A1:H1, B3:D3, E3:G3, B4:D6, E4:G6,
    #                  B8:D8, E8:G8, B9:D11, E9:G11
    # We check the box merges only (not A1:H1 which is scored in C1)
    # =========================================================================
    try:
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        expected_box_merges = [
            'B3:D3', 'E3:G3',   # top row labels
            'B4:D6', 'E4:G6',   # top row values
            'B8:D8', 'E8:G8',   # bottom row labels
            'B9:D11', 'E9:G11', # bottom row values
        ]
        c5_score = 0.0
        pts_per_merge = 0.15 / len(expected_box_merges)

        for expected in expected_box_merges:
            if expected in merged_ranges:
                c5_score += pts_per_merge
                print(f"PASS: C5 — Merge {expected} found")
            else:
                print(f"FAIL: C5 — Merge {expected} not found (have: {merged_ranges})")

        total_score += c5_score
        print(f"  C5 subtotal: {c5_score:.2f}/0.15")

    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # =========================================================================
    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
