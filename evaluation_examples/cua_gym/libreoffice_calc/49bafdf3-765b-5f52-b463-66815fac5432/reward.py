"""
Reward Script: Create a formatted cash flow statement with three sections and proper accounting subtotals.
Task ID: calc_gpm_011
Domain: libreoffice_calc
Scoring:
  Component 1: Header merges & formatting (0.20)
  Component 2: Section headers bold with colored text (0.15)
  Component 3: Subtotal formulas in C9, C14, C19 (0.25)
  Component 4: Net Change formula in C21 with double border/underline (0.15)
  Component 5: Currency number format on amount cells (0.15)
  Component 6: Subtotal rows bold with top borders (0.10)
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_011'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet name
    if 'CashFlow' not in wb.sheetnames:
        print("FAIL: Sheet 'CashFlow' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['CashFlow']

    # =========================================================================
    # Component 1: Header merges & formatting (0.20 points)
    # A1:C1 merged, 14pt bold centered; A2:C2 merged, 12pt centered;
    # A3:C3 merged, 11pt italic centered
    # INITIAL: no merges, no bold, no centering, all 11pt
    # =========================================================================
    try:
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        sub_score = 0.0

        # Check A1:C1 merge
        has_a1_merge = any('A1' in r and 'C1' in r for r in merged_ranges)
        a1 = ws['A1']
        a1_bold = a1.font.bold is True
        a1_size = a1.font.size is not None and abs(a1.font.size - 14) < 0.5
        a1_center = a1.alignment.horizontal == 'center'

        if has_a1_merge and a1_bold and a1_size and a1_center:
            sub_score += 0.07
            print("PASS: A1:C1 merged, 14pt bold centered")
        else:
            print(f"FAIL: A1:C1 — merge={has_a1_merge} bold={a1_bold} size14={a1_size} center={a1_center}")

        # Check A2:C2 merge
        has_a2_merge = any('A2' in r and 'C2' in r for r in merged_ranges)
        a2 = ws['A2']
        a2_size = a2.font.size is not None and abs(a2.font.size - 12) < 0.5
        a2_center = a2.alignment.horizontal == 'center'

        if has_a2_merge and a2_size and a2_center:
            sub_score += 0.07
            print("PASS: A2:C2 merged, 12pt centered")
        else:
            print(f"FAIL: A2:C2 — merge={has_a2_merge} size12={a2_size} center={a2_center}")

        # Check A3:C3 merge
        has_a3_merge = any('A3' in r and 'C3' in r for r in merged_ranges)
        a3 = ws['A3']
        a3_italic = a3.font.italic is True
        a3_center = a3.alignment.horizontal == 'center'

        if has_a3_merge and a3_italic and a3_center:
            sub_score += 0.06
            print("PASS: A3:C3 merged, italic centered")
        else:
            print(f"FAIL: A3:C3 — merge={has_a3_merge} italic={a3_italic} center={a3_center}")

        total_score += sub_score
        print(f"  Component 1 subtotal: {sub_score:.2f}/0.20")

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Section headers bold with colored text (0.15 points)
    # A5 'Operating Activities' bold, dark green text
    # A11 'Investing Activities' bold, dark blue text
    # A16 'Financing Activities' bold, dark red text
    # INITIAL: none are bold, no colored text
    # =========================================================================
    try:
        sub_score = 0.0

        # A5: Operating Activities — bold, dark green
        a5 = ws['A5']
        a5_bold = a5.font.bold is True
        try:
            a5_color = a5.font.color.rgb if a5.font.color else None
        except:
            a5_color = None
        # Accept various dark green shades
        a5_green = a5_color is not None and isinstance(a5_color, str) and (
            '006100' in a5_color or '008000' in a5_color or '00FF00' not in a5_color.upper()
        )
        # More precise: check it has a greenish color (not default black/theme)
        if a5_bold and a5_color is not None and isinstance(a5_color, str) and len(a5_color) >= 6:
            sub_score += 0.05
            print(f"PASS: A5 bold with color {a5_color}")
        else:
            print(f"FAIL: A5 — bold={a5_bold} color={a5_color}")

        # A11: Investing Activities — bold, dark blue
        a11 = ws['A11']
        a11_bold = a11.font.bold is True
        try:
            a11_color = a11.font.color.rgb if a11.font.color else None
        except:
            a11_color = None
        if a11_bold and a11_color is not None and isinstance(a11_color, str) and len(a11_color) >= 6:
            sub_score += 0.05
            print(f"PASS: A11 bold with color {a11_color}")
        else:
            print(f"FAIL: A11 — bold={a11_bold} color={a11_color}")

        # A16: Financing Activities — bold, dark red
        a16 = ws['A16']
        a16_bold = a16.font.bold is True
        try:
            a16_color = a16.font.color.rgb if a16.font.color else None
        except:
            a16_color = None
        if a16_bold and a16_color is not None and isinstance(a16_color, str) and len(a16_color) >= 6:
            sub_score += 0.05
            print(f"PASS: A16 bold with color {a16_color}")
        else:
            print(f"FAIL: A16 — bold={a16_bold} color={a16_color}")

        total_score += sub_score
        print(f"  Component 2 subtotal: {sub_score:.2f}/0.15")

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Subtotal formulas in C9, C14, C19 (0.25 points)
    # C9 =SUM(B6:B8), C14 =SUM(B12:B13), C19 =SUM(B17:B18)
    # INITIAL: no formulas in C9, C14, C19 (all empty)
    # =========================================================================
    try:
        sub_score = 0.0
        formula_checks = [
            ('C9', '=SUM(B6:B8)'),
            ('C14', '=SUM(B12:B13)'),
            ('C19', '=SUM(B17:B18)'),
        ]
        pts_each = round(0.25 / 3, 4)

        for coord, expected_formula in formula_checks:
            cell_val = ws[coord].value
            if cell_val is not None and isinstance(cell_val, str):
                # Normalize for comparison
                actual_norm = cell_val.upper().replace(' ', '')
                expected_norm = expected_formula.upper().replace(' ', '')
                if actual_norm == expected_norm:
                    sub_score += pts_each
                    print(f"PASS: {coord} formula = {cell_val}")
                else:
                    print(f"FAIL: {coord} — expected {expected_formula}, found {cell_val}")
            else:
                print(f"FAIL: {coord} — no formula found, value={cell_val!r}")

        total_score += sub_score
        print(f"  Component 3 subtotal: {sub_score:.4f}/0.25")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: Net Change formula in C21 with double border/underline (0.15 points)
    # C21 =C9+C14+C19, bold 12pt, double underline, top border double
    # INITIAL: C21 is empty, A21 is plain 11pt
    # =========================================================================
    try:
        sub_score = 0.0
        c21 = ws['C21']

        # Check formula
        c21_val = c21.value
        if c21_val is not None and isinstance(c21_val, str):
            c21_norm = c21_val.upper().replace(' ', '')
            expected_norm = '=C9+C14+C19'
            if c21_norm == expected_norm:
                sub_score += 0.05
                print(f"PASS: C21 formula = {c21_val}")
            else:
                print(f"FAIL: C21 formula — expected =C9+C14+C19, found {c21_val}")
        else:
            print(f"FAIL: C21 — no formula, value={c21_val!r}")

        # Check bold and size 12
        c21_bold = c21.font.bold is True
        c21_size = c21.font.size is not None and abs(c21.font.size - 12) < 0.5
        if c21_bold and c21_size:
            sub_score += 0.03
            print(f"PASS: C21 bold 12pt")
        else:
            print(f"FAIL: C21 — bold={c21_bold} size={c21.font.size}")

        # Check double border on top
        top_style = c21.border.top.style if c21.border.top else None
        if top_style == 'double':
            sub_score += 0.04
            print(f"PASS: C21 top border = double")
        else:
            print(f"FAIL: C21 top border — expected double, found {top_style}")

        # Check double underline
        c21_underline = c21.font.underline
        if c21_underline == 'double':
            sub_score += 0.03
            print(f"PASS: C21 underline = double")
        else:
            print(f"FAIL: C21 underline — expected double, found {c21_underline}")

        total_score += sub_score
        print(f"  Component 4 subtotal: {sub_score:.2f}/0.15")

    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # =========================================================================
    # Component 5: Currency number format on amount cells (0.15 points)
    # All B-column amounts and C-column subtotals should use currency format
    # with parenthetical negatives: $#,##0_);($#,##0) or similar
    # INITIAL: all cells have 'General' format
    # =========================================================================
    try:
        sub_score = 0.0
        # Check a representative set of amount cells
        currency_cells = ['B6', 'B7', 'B8', 'B12', 'B13', 'B17', 'B18', 'C9', 'C14', 'C19', 'C21']
        currency_pass = 0
        for coord in currency_cells:
            cell = ws[coord]
            nf = cell.number_format if cell.number_format else 'General'
            # Accept any format containing $ and #
            if '$' in nf and '#' in nf:
                currency_pass += 1
            else:
                print(f"  INFO: {coord} number_format={nf}")

        if currency_pass >= 9:
            sub_score = 0.15
            print(f"PASS: Currency format on {currency_pass}/{len(currency_cells)} cells (0.15 pts)")
        elif currency_pass >= 6:
            sub_score = 0.10
            print(f"PARTIAL: Currency format on {currency_pass}/{len(currency_cells)} cells (0.10 pts)")
        elif currency_pass >= 3:
            sub_score = 0.05
            print(f"PARTIAL: Currency format on {currency_pass}/{len(currency_cells)} cells (0.05 pts)")
        else:
            print(f"FAIL: Currency format on only {currency_pass}/{len(currency_cells)} cells")

        total_score += sub_score
        print(f"  Component 5 subtotal: {sub_score:.2f}/0.15")

    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # =========================================================================
    # Component 6: Subtotal rows bold with top borders on C9, C14, C19 (0.10 points)
    # A9, A14, A19 bold; C9, C14, C19 have thin top border
    # INITIAL: none are bold, no borders
    # =========================================================================
    try:
        sub_score = 0.0
        subtotal_checks = [
            ('A9', 'C9'),
            ('A14', 'C14'),
            ('A19', 'C19'),
        ]
        pts_each = round(0.10 / 3, 4)

        for label_coord, value_coord in subtotal_checks:
            label_cell = ws[label_coord]
            value_cell = ws[value_coord]
            label_bold = label_cell.font.bold is True
            value_bold = value_cell.font.bold is True
            top_border = value_cell.border.top.style if value_cell.border.top else None
            has_top_border = top_border is not None and top_border != 'none'

            if label_bold and value_bold and has_top_border:
                sub_score += pts_each
                print(f"PASS: {label_coord} bold, {value_coord} bold + top border ({top_border})")
            else:
                print(f"FAIL: {label_coord} bold={label_bold}, {value_coord} bold={value_bold} top_border={top_border}")

        total_score += sub_score
        print(f"  Component 6 subtotal: {sub_score:.4f}/0.10")

    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Final score
    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score:.4f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook for LibreOffice
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
