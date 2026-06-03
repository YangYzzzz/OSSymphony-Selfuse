"""
Initial Setup: Cash flow statement with raw data only (no formatting, formulas, or merges)
Task ID: calc_gpm_011
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_011'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'CashFlow'

    # --- Header rows (plain text, no merges, no special formatting) ---
    ws['A1'] = 'Apex Industries'
    ws['A2'] = 'Statement of Cash Flows'
    ws['A3'] = 'Year Ended December 31, 2025'

    # --- Operating Activities section ---
    ws['A5'] = 'Operating Activities'
    ws['A6'] = 'Net Income'
    ws['B6'] = 185000
    ws['A7'] = 'Depreciation'
    ws['B7'] = 42000
    ws['A8'] = 'Changes in Working Capital'
    ws['B8'] = -15000
    ws['A9'] = 'Net Cash from Operations'

    # --- Investing Activities section ---
    ws['A11'] = 'Investing Activities'
    ws['A12'] = 'Purchase of Equipment'
    ws['B12'] = -95000
    ws['A13'] = 'Sale of Investments'
    ws['B13'] = 30000
    ws['A14'] = 'Net Cash from Investing'

    # --- Financing Activities section ---
    ws['A16'] = 'Financing Activities'
    ws['A17'] = 'Loan Proceeds'
    ws['B17'] = 100000
    ws['A18'] = 'Dividend Payments'
    ws['B18'] = -45000
    ws['A19'] = 'Net Cash from Financing'

    # --- Net Change row ---
    ws['A21'] = 'Net Change in Cash'

    # Set reasonable column widths for readability
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
