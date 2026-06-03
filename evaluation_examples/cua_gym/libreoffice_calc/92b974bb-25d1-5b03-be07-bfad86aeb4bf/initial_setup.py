"""
Initial Setup: Sales Territory Regional Summary Task
Task ID: calc_sales_territory_regional_009
Domain: libreoffice_calc

Creates a Transactions sheet with 500 rows of realistic sales data.
No regional summary sheets exist yet — those are the agent's task to create.
"""

import os
import random
import datetime
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_territory_regional_009'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

random.seed(42)

REPS = [
    'Sarah Chen', 'Marcus Johnson', 'Emily Rivera', 'David Kim',
    'Priya Patel', 'James O\'Brien', 'Linda Torres', 'Kevin Zhang',
    'Rachel Adams', 'Michael Scott', 'Aisha Williams', 'Tom Nguyen',
    'Jessica Lee', 'Carlos Mendez', 'Natalie Brooks', 'Omar Hassan',
    'Stephanie Clark', 'Brian Foster', 'Diana Reyes', 'Nathan Park',
]

PRODUCTS = [
    'Enterprise License', 'Professional Suite', 'Starter Pack',
    'Add-on Module', 'Support Contract', 'Training Package',
    'Cloud Subscription', 'On-Premise Install', 'API Access',
    'Analytics Dashboard',
]

REGIONS = ['North', 'South', 'East', 'West']

def random_date(start, end):
    delta = end - start
    return start + datetime.timedelta(days=random.randint(0, delta.days))

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Transactions'

    # --- Header Row ---
    headers = ['Trans ID', 'Date', 'Rep', 'Region', 'Product', 'Revenue', 'Units']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # --- 500 Rows of Realistic Transaction Data ---
    start_date = datetime.date(2024, 1, 1)
    end_date = datetime.date(2025, 3, 31)

    for i in range(1, 501):
        trans_id = f'TXN-{10000 + i}'
        trans_date = random_date(start_date, end_date)
        rep = random.choice(REPS)
        region = random.choice(REGIONS)
        product = random.choice(PRODUCTS)
        revenue = round(random.uniform(2500, 85000), 2)
        units = random.randint(1, 25)

        ws.cell(row=i + 1, column=1, value=trans_id)
        ws.cell(row=i + 1, column=2, value=trans_date.strftime('%Y-%m-%d'))
        ws.cell(row=i + 1, column=3, value=rep)
        ws.cell(row=i + 1, column=4, value=region)
        ws.cell(row=i + 1, column=5, value=product)
        ws.cell(row=i + 1, column=6, value=revenue)
        ws.cell(row=i + 1, column=7, value=units)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheets: {wb.sheetnames}')
    print(f'  Transactions rows: {ws.max_row - 1} (plus header)')

create_initial()
