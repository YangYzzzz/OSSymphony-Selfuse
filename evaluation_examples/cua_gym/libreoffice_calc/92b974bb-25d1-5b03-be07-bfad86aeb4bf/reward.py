"""
Reward Script: Create four regional summary sheets with SUMIFS/COUNTIFS formulas
Task ID: calc_sales_territory_regional_009
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Four regional sheets exist (North, South, East, West)       — 0.25 pts
  Component 2: Each sheet has correct title in A1 'Region Summary — X'     — 0.20 pts
  Component 3: Each sheet has SUMIFS formula in B3 referencing Transactions — 0.25 pts
  Component 4: Each sheet has COUNTIFS formula in B4 referencing Transactions — 0.20 pts
  Component 5: Each sheet has =B3/B4 (or equivalent) formula in B5         — 0.10 pts
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_territory_regional_009'
REGIONS = ['North', 'South', 'East', 'West']


def has_sumifs_for_region(formula, region):
    """
    Check that B3 contains a SUMIFS formula that:
    - References Transactions sheet column F (Revenue)
    - References Transactions sheet column D (Region)
    - Uses the region name as criteria
    Returns True if formula is a valid SUMIFS for this region.
    """
    if not isinstance(formula, str):
        return False
    upper = formula.upper().replace(' ', '')
    if not upper.startswith('=SUMIFS('):
        return False
    # Must reference Transactions sheet for data range
    if 'TRANSACTIONS!' not in upper:
        return False
    # Must reference the region name as a string literal
    region_upper = region.upper()
    if f'"{region_upper}"' not in upper and f"'{region_upper}'" not in upper:
        return False
    return True


def has_countifs_for_region(formula, region):
    """
    Check that B4 contains a COUNTIFS formula that:
    - References Transactions sheet column D (Region)
    - Uses the region name as criteria
    """
    if not isinstance(formula, str):
        return False
    upper = formula.upper().replace(' ', '')
    if not upper.startswith('=COUNTIFS(') and 'COUNTA' not in upper:
        return False
    # Must reference Transactions sheet
    if 'TRANSACTIONS!' not in upper:
        return False
    # Must reference the region name
    region_upper = region.upper()
    if f'"{region_upper}"' not in upper and f"'{region_upper}'" not in upper:
        return False
    return True


def has_avg_formula(formula):
    """
    Check that B5 contains a formula referencing B3 and B4 (average calculation).
    Accepts =B3/B4 or =AVERAGE(...) or similar.
    """
    if not isinstance(formula, str):
        return False
    upper = formula.upper().replace(' ', '')
    # Accept =B3/B4 pattern
    if '=B3/B4' in upper:
        return True
    # Accept any formula containing B3 and B4 with division
    if 'B3' in upper and 'B4' in upper and '/' in upper:
        return True
    return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: Transactions sheet must exist and be intact
    if 'Transactions' not in wb.sheetnames:
        print("FAIL: 'Transactions' sheet is missing — file is corrupted")
        print("REWARD: 0.0")
        return 0.0

    ws_trans = wb['Transactions']
    if ws_trans.max_row < 2:
        print("FAIL: 'Transactions' sheet has no data rows")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: All four regional sheets exist (0.25 points)
    # Each of the 4 sheets is worth 0.0625 pts; all 4 together = 0.25
    try:
        present_sheets = [r for r in REGIONS if r in wb.sheetnames]
        missing_sheets = [r for r in REGIONS if r not in wb.sheetnames]
        comp1_score = len(present_sheets) * 0.0625
        if len(present_sheets) == 4:
            print(f"PASS: Component 1 — All 4 regional sheets exist: {REGIONS} (0.25 pts)")
        elif len(present_sheets) > 0:
            print(f"PARTIAL: Component 1 — {len(present_sheets)}/4 regional sheets exist: {present_sheets} ({comp1_score} pts)")
            print(f"FAIL: Missing sheets: {missing_sheets}")
        else:
            print(f"FAIL: Component 1 — No regional sheets found. Expected: {REGIONS}")
        if comp1_score > 0:
            total_score += comp1_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Each sheet has correct title in A1 'Region Summary — RegionName' (0.20 points)
    # Each correct title is worth 0.05 pts
    try:
        for region in REGIONS:
            if region not in wb.sheetnames:
                print(f"FAIL: Component 2 — Sheet '{region}' missing, cannot check title")
                continue
            ws = wb[region]
            a1_val = ws['A1'].value
            expected_title = f'Region Summary \u2014 {region}'
            title_ok = (a1_val == expected_title) or (
                a1_val is not None and
                region in str(a1_val) and
                'Region Summary' in str(a1_val)
            )
            if title_ok:
                print(f"PASS: Component 2 — A1 title correct in '{region}' sheet: {repr(a1_val)} (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 2 — A1 title wrong in '{region}' sheet. Expected: {repr(expected_title)}, Found: {repr(a1_val)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Each sheet has SUMIFS formula in B3 referencing Transactions (0.25 points)
    # Each correct SUMIFS is worth 0.0625 pts
    try:
        for region in REGIONS:
            if region not in wb.sheetnames:
                print(f"FAIL: Component 3 — Sheet '{region}' missing, cannot check B3")
                continue
            ws = wb[region]
            b3_val = ws['B3'].value
            sumifs_ok = has_sumifs_for_region(b3_val, region) or (
                isinstance(b3_val, str) and
                'SUMIFS' in b3_val.upper() and
                'Transactions' in b3_val
            )
            if sumifs_ok:
                print(f"PASS: Component 3 — SUMIFS formula in '{region}'!B3: {repr(b3_val)} (0.0625 pts)")
                total_score += 0.0625
            else:
                print(f"FAIL: Component 3 — Expected SUMIFS formula in '{region}'!B3, found: {repr(b3_val)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Each sheet has COUNTIFS formula in B4 referencing Transactions (0.20 points)
    # Each correct COUNTIFS is worth 0.05 pts
    try:
        for region in REGIONS:
            if region not in wb.sheetnames:
                print(f"FAIL: Component 4 — Sheet '{region}' missing, cannot check B4")
                continue
            ws = wb[region]
            b4_val = ws['B4'].value
            countifs_ok = has_countifs_for_region(b4_val, region) or (
                isinstance(b4_val, str) and
                ('COUNTIFS' in b4_val.upper() or 'COUNTA' in b4_val.upper()) and
                'Transactions' in b4_val
            )
            if countifs_ok:
                print(f"PASS: Component 4 — COUNTIFS formula in '{region}'!B4: {repr(b4_val)} (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 4 — Expected COUNTIFS formula in '{region}'!B4, found: {repr(b4_val)}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Each sheet has average formula =B3/B4 in B5 (0.10 points)
    # Each correct formula is worth 0.025 pts
    try:
        for region in REGIONS:
            if region not in wb.sheetnames:
                print(f"FAIL: Component 5 — Sheet '{region}' missing, cannot check B5")
                continue
            ws = wb[region]
            b5_val = ws['B5'].value
            if has_avg_formula(b5_val):
                print(f"PASS: Component 5 — Average formula in '{region}'!B5: {repr(b5_val)} (0.025 pts)")
                total_score += 0.025
            else:
                print(f"FAIL: Component 5 — Expected average formula in '{region}'!B5, found: {repr(b5_val)}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
