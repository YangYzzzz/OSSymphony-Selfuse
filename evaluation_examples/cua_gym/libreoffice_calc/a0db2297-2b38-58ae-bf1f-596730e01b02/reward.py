"""
Reward Script: Project Timeline Template in LibreOffice Calc
Task ID: calc_gg2_041
Domain: libreoffice_calc
Scoring:
  Component 1: Month headers B1:M1 (1-12)           — 0.20 pts
  Component 2: Project labels A2:A9 (8 entries)      — 0.15 pts
  Component 3: Start/End headers N1, O1              — 0.10 pts
  Component 4: Column widths B-M set ~9.5 chars      — 0.15 pts
  Component 5: Conditional formatting formula on B2:M9 — 0.30 pts
  Component 6: CF fill is dark blue                  — 0.10 pts
  Total: 1.00
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg2_041'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Timeline' sheet must exist
    if 'Timeline' not in wb.sheetnames:
        print("FAIL: No 'Timeline' sheet found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Timeline']

    # Component 1: Month headers B1:M1 contain values 1-12 (0.20 pts)
    try:
        month_vals = []
        for col_idx in range(2, 14):  # B=2 through M=13
            val = ws.cell(row=1, column=col_idx).value
            month_vals.append(val)

        # Check if values are 1 through 12 (as numbers)
        expected = list(range(1, 13))
        actual_nums = []
        for v in month_vals:
            if v is None:
                actual_nums.append(None)
            else:
                try:
                    actual_nums.append(int(float(v)))
                except (ValueError, TypeError):
                    # Could be month names like "Jan", "Feb", etc.
                    actual_nums.append(v)

        if actual_nums == expected:
            print(f"PASS: Component 1 — Month headers B1:M1 = 1..12 (0.20 pts)")
            total_score += 0.20
        else:
            # Also accept month name abbreviations
            month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                           'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
            month_names_lower = [m.lower() for m in month_names]
            actual_strs = [str(v).strip().lower()[:3] if v else '' for v in month_vals]
            if actual_strs == month_names_lower:
                print(f"PASS: Component 1 — Month headers B1:M1 = Jan..Dec (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 1 — Expected months 1-12 or Jan-Dec, found: {month_vals}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Project labels A2:A9 have 8 non-empty entries (0.15 pts)
    try:
        project_labels = []
        for row_idx in range(2, 10):  # A2:A9
            val = ws.cell(row=row_idx, column=1).value
            project_labels.append(val)

        non_empty = [v for v in project_labels if v is not None and str(v).strip() != '']
        if len(non_empty) >= 8:
            print(f"PASS: Component 2 — 8 project labels in A2:A9 (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 — Expected 8 project labels, found {len(non_empty)}: {project_labels}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Start/End headers in N1 and O1 (0.10 pts)
    try:
        n1_val = ws.cell(row=1, column=14).value  # N1
        o1_val = ws.cell(row=1, column=15).value  # O1

        n1_ok = n1_val is not None and 'start' in str(n1_val).lower()
        o1_ok = o1_val is not None and 'end' in str(o1_val).lower()

        if n1_ok and o1_ok:
            print(f"PASS: Component 3 — N1='{n1_val}', O1='{o1_val}' (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — Expected Start/End headers, found N1={n1_val!r}, O1={o1_val!r}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Column widths B-M set to approximately 9.5 chars (~2.5cm) (0.15 pts)
    # Initial file has no custom widths; golden has all B-M at 9.5
    # Accept range 7.0 to 13.0 (reasonably narrow, not default ~8.43 unless explicitly set)
    try:
        col_letters = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
        widths_set = 0
        for cl in col_letters:
            dim = ws.column_dimensions.get(cl)
            if dim is not None and dim.width is not None:
                w = dim.width
                # Accept widths between 7.0 and 14.0 as intentionally set
                # Default openpyxl width is about 8.43 (or 8.0), so we need
                # to distinguish "set to ~9.5" from "not set at all"
                if 8.5 <= w <= 14.0:
                    widths_set += 1

        if widths_set >= 10:  # At least 10 of 12 columns have appropriate width
            print(f"PASS: Component 4 — {widths_set}/12 columns B-M have width ~9.5 (0.15 pts)")
            total_score += 0.15
        else:
            # Print actual widths for debugging
            actual_widths = {}
            for cl in col_letters:
                dim = ws.column_dimensions.get(cl)
                actual_widths[cl] = dim.width if dim and dim.width else 'default'
            print(f"FAIL: Component 4 — Only {widths_set}/12 columns have width ~9.5. Widths: {actual_widths}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Conditional formatting rule on B2:M9 with correct formula (0.30 pts)
    try:
        cf_list = list(ws.conditional_formatting)
        comp5_score = 0.0

        for cf in cf_list:
            cf_range = str(cf).upper().replace(' ', '')
            for rule in cf.rules:
                if rule.type == 'expression' and rule.formula:
                    formula_str = rule.formula[0].upper().replace(' ', '')
                    if 'AND(' in formula_str and '$N' in formula_str and '$O' in formula_str:
                        # Check range covers B2:M9
                        if 'B2' in cf_range and 'M9' in cf_range:
                            comp5_score = 0.30
                        elif comp5_score < 0.15:
                            comp5_score = 0.15  # formula correct but range mismatch

        if comp5_score >= 0.30:
            print(f"PASS: Component 5 — CF formula rule on B2:M9 with AND/N/O logic (0.30 pts)")
            total_score += 0.30
        elif comp5_score >= 0.15:
            print(f"PARTIAL: Component 5 — CF formula found but range mismatch (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 5 — No conditional formatting with AND formula referencing N/O columns")
            print(f"  CF rules found: {len(cf_list)}")
            for cf in cf_list:
                for rule in cf.rules:
                    print(f"    Range={cf}, type={rule.type}, formula={rule.formula}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: CF fill is dark blue (0.10 pts)
    try:
        cf_list = list(ws.conditional_formatting)
        comp6_color = None

        for cf in cf_list:
            for rule in cf.rules:
                if rule.type == 'expression' and rule.dxf and rule.dxf.fill:
                    fill = rule.dxf.fill
                    if fill.fgColor and fill.fgColor.rgb:
                        fg_rgb = str(fill.fgColor.rgb).upper()
                        # Extract RGB components (skip alpha)
                        rgb_hex = fg_rgb[-6:]  # last 6 chars
                        r = int(rgb_hex[0:2], 16)
                        g = int(rgb_hex[2:4], 16)
                        b = int(rgb_hex[4:6], 16)
                        # Dark blue: blue > red, blue > green, overall dark (sum < 400)
                        if b > r and b >= g and (r + g + b) < 400:
                            comp6_color = fg_rgb

        if comp6_color is not None:
            print(f"PASS: Component 6 — CF fill is dark blue ({comp6_color}) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 6 — CF fill is not dark blue")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    # Round to avoid floating point display issues
    final_score = round(final_score, 2)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
