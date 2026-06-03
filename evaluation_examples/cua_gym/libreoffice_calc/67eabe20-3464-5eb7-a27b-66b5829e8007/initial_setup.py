"""
Initial Setup: Link cell A1 on Finance sheet to external workbook Budget2024.xlsx
Task ID: calc_mcp_047
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_047'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'
EXTERNAL_DIR = f'{WORKDIR}/Documents'
EXTERNAL_FILE = f'{EXTERNAL_DIR}/Budget2024.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_external_workbook():
    """Create the external Budget2024.xlsx that the task references."""
    os.makedirs(EXTERNAL_DIR, exist_ok=True)

    wb = openpyxl.Workbook()

    # --- Annual sheet ---
    ws_annual = wb.active
    ws_annual.title = 'Annual'

    # Headers
    headers = ['Category', 'Budget Amount', 'Actual Amount', 'Variance']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws_annual.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Data rows - B5 must be 250000 (Annual Budget Total)
    budget_data = [
        ['Personnel Costs', 180000, 175420, None],
        ['Office Lease', 48000, 48000, None],
        ['Technology & IT', 35000, 38200, None],
        ['Total Annual Budget', 250000, 248750, None],  # Row 5, B5=250000
        ['Marketing', 42000, 39800, None],
        ['Travel & Conferences', 18000, 15600, None],
        ['Professional Development', 12000, 10950, None],
        ['Miscellaneous', 8000, 7250, None],
    ]
    for r, row_data in enumerate(budget_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_annual.cell(row=r, column=c, value=val)
            if c in (2, 3) and val is not None:
                cell.number_format = '$#,##0'
        # Variance formula in column D
        ws_annual.cell(row=r, column=4, value=f'=C{r}-B{r}')
        ws_annual.cell(row=r, column=4).number_format = '$#,##0'

    # Column widths
    ws_annual.column_dimensions['A'].width = 28
    ws_annual.column_dimensions['B'].width = 16
    ws_annual.column_dimensions['C'].width = 16
    ws_annual.column_dimensions['D'].width = 14

    # --- Q1-Q4 sheet ---
    ws_quarters = wb.create_sheet('Quarterly')
    q_headers = ['Quarter', 'Revenue', 'Expenses', 'Net Income']
    for col, h in enumerate(q_headers, 1):
        cell = ws_quarters.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
    q_data = [
        ['Q1 2024', 125000, 98000, None],
        ['Q2 2024', 138000, 105000, None],
        ['Q3 2024', 142000, 110000, None],
        ['Q4 2024', 155000, 118000, None],
    ]
    for r, row_data in enumerate(q_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_quarters.cell(row=r, column=c, value=val)
        ws_quarters.cell(row=r, column=4, value=f'=B{r}-C{r}')

    wb.save(EXTERNAL_FILE)
    print(f'External workbook created: {EXTERNAL_FILE}')


def create_initial():
    """Create the main task workbook with Finance sheet, A1 empty."""
    wb = openpyxl.Workbook()

    # --- Finance sheet (A1 must be EMPTY - task will fill it) ---
    ws_finance = wb.active
    ws_finance.title = 'Finance'

    # Headers starting at row 2 (A1 is reserved for the external link)
    # Row 1 label in B1 onwards
    ws_finance['B1'] = 'Department Financial Overview'
    ws_finance['B1'].font = Font(name='Calibri', size=14, bold=True, color='1F4E79')

    # Data headers in row 3
    fin_headers = ['Linked Budget', 'Department', 'Q1 Spend', 'Q2 Spend', 'Q3 Spend', 'Q4 Spend', 'Annual Total']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for col, h in enumerate(fin_headers, 1):
        cell = ws_finance.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Financial data rows
    fin_data = [
        [None, 'Engineering', 42500, 44800, 46200, 48100, None],
        [None, 'Marketing', 18200, 19500, 17800, 21000, None],
        [None, 'Sales', 31000, 33500, 35200, 37800, None],
        [None, 'Operations', 22000, 21500, 23000, 24500, None],
        [None, 'Human Resources', 15800, 16200, 15900, 16500, None],
        [None, 'Research & Development', 38000, 41000, 39500, 42000, None],
        [None, 'Customer Support', 12500, 13000, 12800, 13200, None],
        [None, 'Legal & Compliance', 8500, 9000, 8800, 9200, None],
        [None, 'Finance & Accounting', 11000, 11500, 11200, 11800, None],
        [None, 'Executive Office', 25000, 25000, 25000, 25000, None],
    ]
    for r, row_data in enumerate(fin_data, 4):
        for c, val in enumerate(row_data, 1):
            cell = ws_finance.cell(row=r, column=c, value=val)
            if c >= 3 and val is not None:
                cell.number_format = '$#,##0'
            cell.border = thin_border
        # Annual total formula
        ws_finance.cell(row=r, column=7, value=f'=SUM(C{r}:F{r})')
        ws_finance.cell(row=r, column=7).number_format = '$#,##0'

    # Note: A1 is intentionally left EMPTY for the task
    ws_finance.column_dimensions['A'].width = 18
    ws_finance.column_dimensions['B'].width = 24
    for col_letter in ['C', 'D', 'E', 'F', 'G']:
        ws_finance.column_dimensions[col_letter].width = 14

    # --- Revenue sheet ---
    ws_revenue = wb.create_sheet('Revenue')
    rev_headers = ['Month', 'Product Sales', 'Service Revenue', 'Subscriptions', 'Total']
    for col, h in enumerate(rev_headers, 1):
        cell = ws_revenue.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FF548235', end_color='FF548235', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')

    months = ['January', 'February', 'March', 'April', 'May', 'June',
              'July', 'August', 'September', 'October', 'November', 'December']
    rev_data = [
        [45200, 22100, 8500], [47800, 23400, 8800], [51000, 24800, 9100],
        [48500, 25200, 9400], [52300, 26100, 9700], [55000, 27500, 10000],
        [53200, 28000, 10300], [56800, 29200, 10600], [58100, 30500, 10900],
        [61000, 31800, 11200], [63500, 33000, 11500], [67000, 34500, 11800],
    ]
    for r, (month, data) in enumerate(zip(months, rev_data), 2):
        ws_revenue.cell(row=r, column=1, value=month)
        for c, val in enumerate(data, 2):
            ws_revenue.cell(row=r, column=c, value=val)
            ws_revenue.cell(row=r, column=c).number_format = '$#,##0'
        ws_revenue.cell(row=r, column=5, value=f'=SUM(B{r}:D{r})')
        ws_revenue.cell(row=r, column=5).number_format = '$#,##0'

    # --- Expenses sheet ---
    ws_expenses = wb.create_sheet('Expenses')
    exp_headers = ['Category', 'Jan-Mar', 'Apr-Jun', 'Jul-Sep', 'Oct-Dec', 'Annual']
    for col, h in enumerate(exp_headers, 1):
        cell = ws_expenses.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='FFC00000', end_color='FFC00000', fill_type='solid')

    exp_data = [
        ['Salaries & Wages', 145000, 148000, 151000, 154000],
        ['Benefits & Insurance', 42000, 42500, 43000, 43500],
        ['Rent & Utilities', 36000, 36000, 36000, 36000],
        ['Software Licenses', 15000, 15500, 16000, 16500],
        ['Office Supplies', 4200, 3800, 4500, 5100],
        ['Travel', 8500, 12000, 9500, 7000],
        ['Professional Services', 22000, 18000, 25000, 20000],
        ['Equipment', 15000, 8000, 12000, 18000],
    ]
    for r, row_data in enumerate(exp_data, 2):
        ws_expenses.cell(row=r, column=1, value=row_data[0])
        for c, val in enumerate(row_data[1:], 2):
            ws_expenses.cell(row=r, column=c, value=val)
            ws_expenses.cell(row=r, column=c).number_format = '$#,##0'
        ws_expenses.cell(row=r, column=6, value=f'=SUM(B{r}:E{r})')
        ws_expenses.cell(row=r, column=6).number_format = '$#,##0'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


# Execute
create_external_workbook()
create_initial()
