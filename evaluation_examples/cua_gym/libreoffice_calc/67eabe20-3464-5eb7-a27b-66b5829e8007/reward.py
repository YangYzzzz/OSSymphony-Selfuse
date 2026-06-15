"""
Reward Script: External workbook reference in Finance!A1
Task ID: calc_mcp_047
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): A1 contains a formula (starts with '=')
  Component 2 (0.3): Formula references Budget2024.xlsx file path
  Component 3 (0.3): Formula references Annual sheet cell B5
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_047'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Finance sheet must exist
    if 'Finance' not in wb.sheetnames:
        print("FAIL: 'Finance' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Finance']
    a1_value = ws['A1'].value

    print(f"DEBUG: Finance!A1 raw value = {repr(a1_value)}")

    # Component 1: A1 contains a formula (0.4 points)
    # In initial_env, A1 is None/empty. In golden_env, it should be a formula string.
    try:
        if a1_value is not None and isinstance(a1_value, str) and a1_value.strip().startswith('='):
            print(f"PASS: Component 1 -- A1 contains a formula (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 -- A1 does not contain a formula, found: {repr(a1_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Formula references Budget2024.xlsx (0.3 points)
    # The formula should contain the path to the external file
    try:
        if a1_value is not None and isinstance(a1_value, str):
            formula_upper = a1_value.upper()
            if 'BUDGET2024.XLSX' in formula_upper:
                print(f"PASS: Component 2 -- Formula references Budget2024.xlsx (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 -- Formula does not reference Budget2024.xlsx, found: {repr(a1_value)}")
        else:
            print(f"FAIL: Component 2 -- A1 is not a string formula, found: {repr(a1_value)}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Formula references Annual sheet cell B5 (0.3 points)
    # LibreOffice uses dot notation: Annual.B5 or could also be Annual!B5 or 'Annual'.B5
    try:
        if a1_value is not None and isinstance(a1_value, str):
            formula_upper = a1_value.upper()
            # Check for Annual.B5 or Annual!B5 or 'Annual'.B5 patterns
            if re.search(r"ANNUAL[.'!]B5", formula_upper) or re.search(r"'ANNUAL'\.B5", formula_upper):
                print(f"PASS: Component 3 -- Formula references Annual.B5 (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 3 -- Formula does not reference Annual.B5, found: {repr(a1_value)}")
        else:
            print(f"FAIL: Component 3 -- A1 is not a string formula, found: {repr(a1_value)}")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
