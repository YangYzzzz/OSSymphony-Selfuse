"""
Initial Setup: Library book tracking system
Task ID: calc_wf_051
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_051'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()

    # ====== CATALOG SHEET ======
    ws_cat = wb.active
    ws_cat.title = 'Catalog'

    cat_headers = ['Title', 'Author', 'Genre', 'ISBN', 'Copies']
    header_font = Font(bold=True)
    for col, h in enumerate(cat_headers, 1):
        cell = ws_cat.cell(row=1, column=col, value=h)
        cell.font = header_font

    authors = [
        'Margaret Atwood', 'Haruki Murakami', 'Chimamanda Ngozi Adichie',
        'Gabriel Garcia Marquez', 'Toni Morrison', 'Jorge Luis Borges',
        'Ursula K. Le Guin', 'Kazuo Ishiguro', 'Octavia Butler',
        'Elena Ferrante', 'Salman Rushdie', 'Isabel Allende',
        'Yukio Mishima', 'Jhumpa Lahiri', 'Roberto Bolano',
        'Arundhati Roy', 'Orhan Pamuk', 'Alice Munro',
        'Patrick Modiano', 'Olga Tokarczuk',
    ]

    titles = [
        'The Blind Assassin', 'Norwegian Wood', 'Half of a Yellow Sun',
        'One Hundred Years of Solitude', 'Beloved', 'Ficciones',
        'The Left Hand of Darkness', 'Never Let Me Go', 'Kindred',
        'My Brilliant Friend', 'Midnight\'s Children', 'The House of the Spirits',
        'The Temple of the Golden Pavilion', 'Interpreter of Maladies', '2666',
        'The God of Small Things', 'My Name Is Red', 'Dear Life',
        'Missing Person', 'Flights', 'The Testaments', 'Kafka on the Shore',
        'Americanah', 'Love in the Time of Cholera', 'Song of Solomon',
        'The Aleph', 'A Wizard of Earthsea', 'The Remains of the Day',
        'Parable of the Sower', 'The Story of a New Name',
        'Shame', 'Eva Luna', 'Confessions of a Mask', 'The Namesake',
        'The Savage Detectives', 'An Atlas of Impossible Longing',
        'Snow', 'Runaway', 'Dora Bruder', 'Drive Your Plow Over the Bones of the Dead',
        'The Handmaid\'s Testaments', 'Wind-Up Bird Chronicle',
        'Purple Hibiscus', 'Chronicle of a Death Foretold', 'Jazz',
        'The Library of Babel', 'The Dispossessed', 'An Artist of the Floating World',
        'Wild Seed', 'Those Who Leave and Those Who Stay',
    ]

    genres = ['Fiction', 'Literary Fiction', 'Science Fiction', 'Fantasy',
              'Historical Fiction', 'Mystery', 'Biography', 'Romance',
              'Thriller', 'Poetry']

    books = []
    for i in range(50):
        isbn = f'978-{random.randint(0,9)}-{random.randint(1000,9999)}-{random.randint(1000,9999)}-{random.randint(0,9)}'
        copies = random.randint(1, 5)
        author = authors[i % len(authors)]
        title = titles[i] if i < len(titles) else f'{titles[i % len(titles)]} Vol. {i // len(titles) + 1}'
        genre = genres[i % len(genres)]
        books.append([title, author, genre, isbn, copies])

    for r, book in enumerate(books, 2):
        for c, val in enumerate(book, 1):
            ws_cat.cell(row=r, column=c, value=val)

    # Column widths
    ws_cat.column_dimensions['A'].width = 40
    ws_cat.column_dimensions['B'].width = 28
    ws_cat.column_dimensions['C'].width = 18
    ws_cat.column_dimensions['D'].width = 24
    ws_cat.column_dimensions['E'].width = 10

    # ====== LOANS SHEET ======
    ws_loans = wb.create_sheet('Loans')

    loan_headers = ['Book ISBN', 'Borrower', 'Checkout Date', 'Due Date', 'Return Date']
    for col, h in enumerate(loan_headers, 1):
        cell = ws_loans.cell(row=1, column=col, value=h)
        cell.font = header_font

    borrowers = [
        'Emily Rodriguez', 'James Okafor', 'Priya Sharma', 'David Kim',
        'Sarah Mitchell', 'Wei Zhang', 'Ana Petrova', 'Lucas Fernandez',
        'Fatima Al-Rashid', 'Oliver Bennett', 'Yuki Tanaka', 'Maria Costa',
        'Hassan Diop', 'Clara Johansson', 'Raj Patel', 'Sophie Laurent',
        'Tomasz Kowalski', 'Lena Muller', 'Carlos Mendoza', 'Ingrid Svensson',
    ]

    # Generate 80 loan records
    # Make some still out (no return date) and some overdue
    base_date = datetime(2025, 1, 1)
    loans = []
    for i in range(80):
        isbn = books[i % 50][3]  # cycle through book ISBNs
        borrower = borrowers[i % len(borrowers)]
        # Spread checkout dates over 10 months
        checkout_offset = random.randint(0, 300)
        checkout_date = base_date + timedelta(days=checkout_offset)
        due_date = checkout_date + timedelta(days=14)

        # About 15 loans still out (no return date)
        if i < 15:
            return_date = None  # still out
        else:
            # Returned: return within 1-21 days of checkout
            return_offset = random.randint(1, 21)
            return_date = checkout_date + timedelta(days=return_offset)

        loans.append([isbn, borrower, checkout_date, due_date, return_date])

    for r, loan in enumerate(loans, 2):
        for c, val in enumerate(loan, 1):
            cell = ws_loans.cell(row=r, column=c, value=val)
            if c in (3, 4, 5) and val is not None:
                cell.number_format = 'yyyy-mm-dd'

    ws_loans.column_dimensions['A'].width = 24
    ws_loans.column_dimensions['B'].width = 22
    ws_loans.column_dimensions['C'].width = 16
    ws_loans.column_dimensions['D'].width = 16
    ws_loans.column_dimensions['E'].width = 16

    # ====== DASHBOARD SHEET ======
    ws_dash = wb.create_sheet('Dashboard')
    ws_dash['A1'] = 'Library Dashboard'
    ws_dash['A1'].font = Font(size=16, bold=True)

    # Labels only - NO formulas, NO computed values
    ws_dash['A3'] = 'Currently Loaned Books:'
    ws_dash['A4'] = 'Overdue Items (14+ days):'
    ws_dash['A5'] = 'Most Popular Book (ISBN):'
    ws_dash['A6'] = 'Average Loan Duration (days):'

    ws_dash['A3'].font = Font(bold=True)
    ws_dash['A4'].font = Font(bold=True)
    ws_dash['A5'].font = Font(bold=True)
    ws_dash['A6'].font = Font(bold=True)

    ws_dash.column_dimensions['A'].width = 32
    ws_dash.column_dimensions['B'].width = 20

    # Monthly checkout summary labels (for chart later)
    ws_dash['A8'] = 'Monthly Checkout Summary'
    ws_dash['A8'].font = Font(size=12, bold=True)
    ws_dash['A9'] = 'Month'
    ws_dash['B9'] = 'Checkouts'
    ws_dash['A9'].font = Font(bold=True)
    ws_dash['B9'].font = Font(bold=True)

    months = ['Jan 2025', 'Feb 2025', 'Mar 2025', 'Apr 2025', 'May 2025',
              'Jun 2025', 'Jul 2025', 'Aug 2025', 'Sep 2025', 'Oct 2025']
    for i, m in enumerate(months):
        ws_dash.cell(row=10 + i, column=1, value=m)

    # Count checkouts per month from actual loan data
    month_counts = {}
    for loan in loans:
        checkout = loan[2]
        key = checkout.strftime('%b %Y')
        month_counts[key] = month_counts.get(key, 0) + 1

    for i, m in enumerate(months):
        ws_dash.cell(row=10 + i, column=2, value=month_counts.get(m, 0))

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
