"""
Initial Setup: Multi-month expense report workbook (blank sheets)
Task ID: calc_fin_expense_report_multisheet_032
Domain: libreoffice_calc

Creates a workbook with 4 blank sheets: January, February, March, Q1_Summary.
The task is to set up headers, formulas, and a Q1 rollup from scratch.
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_expense_report_multisheet_032'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # Sheet 1: January (blank)
    ws_jan = wb.active
    ws_jan.title = 'January'

    # Sheet 2: February (blank)
    ws_feb = wb.create_sheet('February')

    # Sheet 3: March (blank)
    ws_mar = wb.create_sheet('March')

    # Sheet 4: Q1_Summary (blank)
    ws_q1 = wb.create_sheet('Q1_Summary')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets created (all blank): January, February, March, Q1_Summary')


create_initial()
