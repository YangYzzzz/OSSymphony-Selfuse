"""
Reward Script: Multi-month expense report workbook with Q1 rollup
Task ID: calc_fin_expense_report_multisheet_032
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.25): Month sheets have correct headers, bold row 1, freeze_panes=A2
  - Component 2 (0.25): Month sheets have D column currency format and SUM formula in D31 with label
  - Component 3 (0.30): Q1_Summary sheet has correct structure, cross-sheet formulas, bold totals
  - Component 4 (0.20): Bar chart on Q1_Summary sheet
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_expense_report_multisheet_032'

MONTH_SHEETS = ['January', 'February', 'March']
EXPECTED_HEADERS = ['Date', 'Employee', 'Category', 'Amount', 'Approved']


def check_month_headers_and_freeze(ws, sheet_name):
    """Check if month sheet has correct headers (bold, row 1) and freeze_panes=A2."""
    actual_headers = [ws.cell(row=1, column=c).value for c in range(1, 6)]
    headers_correct = actual_headers == EXPECTED_HEADERS
    headers_bold = all(ws.cell(row=1, column=c).font.bold == True for c in range(1, 6))
    freeze_ok = (ws.freeze_panes == 'A2')
    return headers_correct and headers_bold and freeze_ok, {
        'headers_correct': headers_correct,
        'headers_bold': headers_bold,
        'freeze_ok': freeze_ok,
        'actual_headers': actual_headers,
    }


def check_month_d_column(ws, sheet_name):
    """Check D2 currency format, D31 SUM formula, C31 'Monthly Total' (bold)."""
    d2_format = ws.cell(row=2, column=4).number_format
    d2_currency = '$#,##0.00' in d2_format or '#,##0.00' in d2_format
    d31_value = ws.cell(row=31, column=4).value
    d31_formula = (
        isinstance(d31_value, str) and
        'SUM' in d31_value.upper() and
        'D2' in d31_value.upper() and
        'D30' in d31_value.upper()
    )
    d31_format = ws.cell(row=31, column=4).number_format
    d31_currency = '$#,##0.00' in d31_format or '#,##0.00' in d31_format
    c31_value = ws.cell(row=31, column=3).value
    c31_label = isinstance(c31_value, str) and 'monthly total' in c31_value.lower()
    c31_bold = ws.cell(row=31, column=3).font.bold == True
    passed = d2_currency and d31_formula and d31_currency and c31_label and c31_bold
    return passed, {
        'd2_currency': d2_currency,
        'd31_formula': d31_formula,
        'd31_currency': d31_currency,
        'c31_label': c31_label,
        'c31_bold': c31_bold,
        'd2_format': d2_format,
        'd31_value': d31_value,
        'c31_value': c31_value,
    }


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # --- Precondition gate: required sheets must exist ---
    required_sheets = MONTH_SHEETS + ['Q1_Summary']
    for sheet_name in required_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"CRITICAL: Required sheet '{sheet_name}' not found in workbook.")
            print(f"REWARD: 0.0")
            return 0.0

    # --- Component 1: Month sheets have correct headers (bold), freeze_panes=A2 (0.25 pts) ---
    # This FAILS on initial file (blank sheets) and PASSES on golden file
    try:
        comp1_pass_count = 0
        for sheet_name in MONTH_SHEETS:
            ws = wb[sheet_name]
            passed, info = check_month_headers_and_freeze(ws, sheet_name)
            if passed:
                comp1_pass_count += 1
                print(f"PASS: {sheet_name} — headers correct, bold, freeze_panes=A2")
            else:
                reasons = []
                if not info['headers_correct']:
                    reasons.append(f"headers: expected {EXPECTED_HEADERS}, got {info['actual_headers']}")
                if not info['headers_bold']:
                    reasons.append("row 1 not bold")
                if not info['freeze_ok']:
                    reasons.append(f"freeze_panes={wb[sheet_name].freeze_panes} (expected A2)")
                print(f"FAIL: {sheet_name} — {'; '.join(reasons)}")

        if comp1_pass_count == 3:
            print(f"PASS: Component 1 — all 3 month sheets have correct headers, bold, freeze_panes=A2 (0.25 pts)")
            total_score += 0.25
        elif comp1_pass_count == 2:
            print(f"PARTIAL: Component 1 — 2/3 month sheets correct (0.16 pts)")
            total_score += 0.16
        elif comp1_pass_count == 1:
            print(f"PARTIAL: Component 1 — 1/3 month sheets correct (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 1 — no month sheets have correct headers/bold/freeze setup")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # --- Component 2: Month sheets D column currency + SUM formula in D31 + label (0.25 pts) ---
    # This FAILS on initial file (blank sheets) and PASSES on golden file
    try:
        comp2_pass_count = 0
        for sheet_name in MONTH_SHEETS:
            ws = wb[sheet_name]
            passed, info = check_month_d_column(ws, sheet_name)
            if passed:
                comp2_pass_count += 1
                print(f"PASS: {sheet_name} — D col currency, D31 SUM formula, C31 'Monthly Total' (bold)")
            else:
                reasons = []
                if not info['d2_currency']:
                    reasons.append(f"D2 number_format={info['d2_format']} (expected currency)")
                if not info['d31_formula']:
                    reasons.append(f"D31 value={repr(info['d31_value'])} (expected =SUM(D2:D30))")
                if not info['d31_currency']:
                    reasons.append(f"D31 format={wb[sheet_name].cell(row=31, column=4).number_format} (expected currency)")
                if not info['c31_label']:
                    reasons.append(f"C31={repr(info['c31_value'])} (expected 'Monthly Total')")
                if not info['c31_bold']:
                    reasons.append("C31 not bold")
                print(f"FAIL: {sheet_name} — {'; '.join(reasons)}")

        if comp2_pass_count == 3:
            print(f"PASS: Component 2 — all 3 month sheets have currency format and SUM in D31 (0.25 pts)")
            total_score += 0.25
        elif comp2_pass_count == 2:
            print(f"PARTIAL: Component 2 — 2/3 month sheets correct (0.16 pts)")
            total_score += 0.16
        elif comp2_pass_count == 1:
            print(f"PARTIAL: Component 2 — 1/3 month sheets correct (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 2 — no month sheets have correct D-column setup")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # --- Component 3: Q1_Summary sheet structure with cross-sheet formulas and bold totals (0.30 pts) ---
    # This FAILS on initial file (blank Q1_Summary) and PASSES on golden file
    try:
        ws_q1 = wb['Q1_Summary']
        sub_score = 0.0

        # Sub-check 3a: Headers in row 1 (Month, Total Expenses, Avg Per Day)
        expected_q1_headers = ['Month', 'Total Expenses', 'Avg Per Day']
        actual_q1_headers = [ws_q1.cell(row=1, column=c).value for c in range(1, 4)]
        if actual_q1_headers == expected_q1_headers:
            print(f"PASS: Q1_Summary headers correct: {actual_q1_headers}")
            sub_score += 0.08
        else:
            print(f"FAIL: Q1_Summary headers: expected {expected_q1_headers}, got {actual_q1_headers}")

        # Sub-check 3b: Month labels in A2:A4 (check all 3 match)
        month_labels_pass_count = sum(
            1 for i, month in enumerate(MONTH_SHEETS, 2)
            if ws_q1.cell(row=i, column=1).value == month
        )
        for i, month in enumerate(MONTH_SHEETS, 2):
            a_val = ws_q1.cell(row=i, column=1).value
            if a_val != month:
                print(f"FAIL: Q1_Summary A{i}: expected '{month}', got {repr(a_val)}")
        if month_labels_pass_count == 3:
            print(f"PASS: Q1_Summary A2:A4 month labels correct")
            sub_score += 0.06
        else:
            print(f"FAIL: Q1_Summary A2:A4 only {month_labels_pass_count}/3 month labels correct")

        # Sub-check 3c: Cross-sheet references in B2:B4 (formula referencing each month's D31)
        cross_refs_pass_count = 0
        for i, month in enumerate(MONTH_SHEETS, 2):
            b_val = ws_q1.cell(row=i, column=2).value
            if isinstance(b_val, str) and month in b_val and 'D31' in b_val:
                cross_refs_pass_count += 1
            else:
                print(f"FAIL: Q1_Summary B{i}: expected formula referencing {month}!D31, got {repr(b_val)}")
        if cross_refs_pass_count == 3:
            print(f"PASS: Q1_Summary B2:B4 cross-sheet formula references correct")
            sub_score += 0.06

        # Sub-check 3d: Q1 Total row (A5='Q1 Total', B5=SUM formula, both bold)
        a5_val = ws_q1.cell(row=5, column=1).value
        b5_val = ws_q1.cell(row=5, column=2).value
        a5_bold = ws_q1.cell(row=5, column=1).font.bold
        b5_bold = ws_q1.cell(row=5, column=2).font.bold
        a5_ok = isinstance(a5_val, str) and 'q1 total' in a5_val.lower()
        b5_formula_ok = (
            isinstance(b5_val, str) and
            'SUM' in b5_val.upper() and
            'B2' in b5_val and
            'B4' in b5_val
        )
        if a5_ok and b5_formula_ok and a5_bold and b5_bold:
            print(f"PASS: Q1_Summary A5='Q1 Total' (bold), B5=SUM(B2:B4) (bold)")
            sub_score += 0.05
        else:
            reasons = []
            if not a5_ok:
                reasons.append(f"A5={repr(a5_val)}")
            if not b5_formula_ok:
                reasons.append(f"B5={repr(b5_val)}")
            if not (a5_bold and b5_bold):
                reasons.append(f"A5 bold={a5_bold}, B5 bold={b5_bold}")
            print(f"FAIL: Q1_Summary Q1 Total row — {'; '.join(reasons)}")

        # Sub-check 3e: Avg Per Day formulas in C2:C4
        avg_pass_count = 0
        for row, divisor in [(2, 31), (3, 28), (4, 31)]:
            c_val = ws_q1.cell(row=row, column=3).value
            if isinstance(c_val, str) and c_val.startswith('=') and f'B{row}' in c_val:
                avg_pass_count += 1
            else:
                print(f"FAIL: Q1_Summary C{row}: expected avg formula like =B{row}/{divisor}, got {repr(c_val)}")
        if avg_pass_count == 3:
            print(f"PASS: Q1_Summary C2:C4 avg per day formulas present")
            sub_score += 0.05

        component3_score = min(sub_score, 0.30)
        if component3_score > 0:
            total_score += component3_score
        if component3_score >= 0.30:
            print(f"PASS: Component 3 — Q1_Summary structure fully correct (0.30 pts)")
        else:
            print(f"PARTIAL: Component 3 — Q1_Summary partial score ({component3_score:.2f}/0.30 pts)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # --- Component 4: Bar chart on Q1_Summary sheet (0.20 pts) ---
    # This FAILS on initial file (blank Q1_Summary) and PASSES on golden file
    try:
        ws_q1 = wb['Q1_Summary']
        charts = ws_q1._charts
        num_charts = len(charts)
        if num_charts >= 1:
            chart = charts[0]
            chart_type = chart.__class__.__name__
            # Check it's a bar chart — task requires "Bar chart showing B2:B4 by month"
            if 'BarChart' in chart_type:
                print(f"PASS: Component 4 — BarChart found on Q1_Summary (0.20 pts)")
                total_score += 0.20
            elif chart_type:
                print(f"PARTIAL: Component 4 — Chart found but type={chart_type}, expected BarChart (0.10 pts)")
                total_score += 0.10
        else:
            print(f"FAIL: Component 4 — No chart found on Q1_Summary sheet")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
