"""
Reward Script: Apply Accounting number format with USD ($) to expense amounts
Task ID: calc_gg5_003
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): All cells C2:C20 have Accounting number format
  Component 2 (0.3): Format includes $ symbol, comma separators, and 2 decimal places
  Component 3 (0.2): Data values in C2:C20 are preserved (not corrupted by format change)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_003'

# The exact Accounting format string that LibreOffice/Excel uses for USD
ACCOUNTING_FORMAT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

# Known valid Accounting-style format patterns (with $ and 2 decimals, parentheses for negatives)
# Different apps may produce slightly different variants
ACCOUNTING_KEYWORDS = ['$', '#,##0.00', '(']

# Expected values in C2:C20 (must be preserved after formatting)
EXPECTED_VALUES = {
    2: 45230, 3: 28750.5, 4: 15000, 5: 62400.75, 6: 8500.25,
    7: 32000, 8: 19875.6, 9: 11200, 10: 54300.8, 11: 37650,
    12: 22100.45, 13: 9800.3, 14: 6750, 15: 14500.9, 16: -3200.5,
    17: 17850, 18: 41000.25, 19: 26300, 20: -1500.75,
}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Q1' sheet must exist
    if 'Q1' not in wb.sheetnames:
        print(f"CRITICAL: 'Q1' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Q1']

    # Component 1: All cells C2:C20 have the Accounting number format (0.5 points)
    # This is the PRIMARY task requirement - changing from plain Number to Accounting
    try:
        accounting_count = 0
        total_cells = 19  # C2 through C20
        for r in range(2, 21):
            cell = ws.cell(row=r, column=3)
            fmt = cell.number_format
            if fmt == ACCOUNTING_FORMAT:
                accounting_count += 1
            else:
                print(f"  C{r}: format={fmt!r} (not exact Accounting match)")

        if accounting_count == total_cells:
            print(f"PASS: Component 1 — All 19 cells have exact Accounting format (0.5 pts)")
            total_score += 0.5
        elif accounting_count > 0:
            # Partial credit: proportional to how many cells have the format
            partial = 0.5 * (accounting_count / total_cells)
            print(f"PARTIAL: Component 1 — {accounting_count}/{total_cells} cells have Accounting format ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No cells have Accounting format. Sample C2: {ws.cell(row=2, column=3).number_format!r}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Format includes $ symbol, comma separators, and 2 decimal places (0.3 points)
    # Checks that even if not an exact match, the format is functionally Accounting-like
    # This distinguishes from initial state where format is '#,##0.00' (no $ sign)
    try:
        accounting_like_count = 0
        for r in range(2, 21):
            cell = ws.cell(row=r, column=3)
            fmt = cell.number_format
            # Must contain $, comma grouping with 2 decimals, and parentheses for negatives
            has_dollar = '$' in fmt
            has_decimal_format = '#,##0.00' in fmt
            has_parens = '(' in fmt and ')' in fmt  # Accounting uses parentheses for negatives
            if has_dollar and has_decimal_format and has_parens:
                accounting_like_count += 1

        if accounting_like_count == total_cells:
            print(f"PASS: Component 2 — All 19 cells have $ + comma + 2 decimals + parentheses (0.3 pts)")
            total_score += 0.3
        elif accounting_like_count > 0:
            partial = 0.3 * (accounting_like_count / total_cells)
            print(f"PARTIAL: Component 2 — {accounting_like_count}/{total_cells} cells have Accounting characteristics ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No cells have Accounting-like format. Sample C2: {ws.cell(row=2, column=3).number_format!r}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Data values are preserved after format change (0.2 points)
    # Only scores if format was actually changed (prevents scoring initial state)
    try:
        # Gate: at least some cells must have a non-initial format to earn these points
        changed_formats = [r for r in range(2, 21)
                          if ws.cell(row=r, column=3).number_format != '#,##0.00']

        if len(changed_formats) == 0:
            print(f"FAIL: Component 3 — Format unchanged from initial; no points for data preservation")
        else:
            preserved_count = 0
            for r in range(2, 21):
                cell = ws.cell(row=r, column=3)
                expected = EXPECTED_VALUES.get(r)
                if expected is not None and cell.value is not None:
                    try:
                        if abs(float(cell.value) - expected) < 0.01:
                            preserved_count += 1
                        else:
                            print(f"  C{r}: expected {expected}, found {cell.value}")
                    except (ValueError, TypeError):
                        print(f"  C{r}: cannot compare, value={cell.value!r}")

            if preserved_count == total_cells:
                print(f"PASS: Component 3 — All 19 data values preserved after format change (0.2 pts)")
                total_score += 0.2
            elif preserved_count > 0:
                partial = 0.2 * (preserved_count / total_cells)
                print(f"PARTIAL: Component 3 — {preserved_count}/{total_cells} values preserved ({partial:.2f} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 3 — No data values preserved")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point: persist app state then verify
def persist_app_state(domain):
    import os, time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
