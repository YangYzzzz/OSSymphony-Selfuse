"""
Initial Setup: Apply Accounting number format to expense amounts in budget spreadsheet
Task ID: calc_gg5_003
Domain: libreoffice_calc

Creates budget.xlsx with Q1 sheet containing department expenses as raw numbers
(no currency formatting). Also includes a Q2 sheet for realism.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_003'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Q1 Sheet ---
    ws1 = wb.active
    ws1.title = 'Q1'

    # Headers
    headers = ['ID', 'Department', 'Expense Amount']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FFD9E2F3', end_color='FFD9E2F3', fill_type='solid')
    header_alignment = Alignment(horizontal='center')
    thin_border = Border(
        bottom=Side(style='thin', color='000000')
    )

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Department expense data - 19 rows (rows 2-20)
    # Realistic budget data with varied amounts, including a few negative values
    data = [
        [1, 'Engineering', 45230.00],
        [2, 'Marketing', 28750.50],
        [3, 'Human Resources', 15000.00],
        [4, 'Sales', 62400.75],
        [5, 'Finance', 8500.25],
        [6, 'Legal', 32000.00],
        [7, 'Operations', 19875.60],
        [8, 'Customer Support', 11200.00],
        [9, 'Research & Development', 54300.80],
        [10, 'IT Infrastructure', 37650.00],
        [11, 'Executive Office', 22100.45],
        [12, 'Facilities', 9800.30],
        [13, 'Training & Development', 6750.00],
        [14, 'Quality Assurance', 14500.90],
        [15, 'Procurement', -3200.50],
        [16, 'Communications', 17850.00],
        [17, 'Product Management', 41000.25],
        [18, 'Data Analytics', 26300.00],
        [19, 'Supply Chain', -1500.75],
    ]

    for r, row_data in enumerate(data, 2):
        ws1.cell(row=r, column=1, value=row_data[0])
        ws1.cell(row=r, column=2, value=row_data[1])
        cell_c = ws1.cell(row=r, column=3, value=row_data[2])
        # Format as plain number with 2 decimals (NOT accounting/currency)
        cell_c.number_format = '#,##0.00'

    # Set column widths
    ws1.column_dimensions['A'].width = 6
    ws1.column_dimensions['B'].width = 28
    ws1.column_dimensions['C'].width = 18

    # Freeze header row
    ws1.freeze_panes = 'A2'

    # --- Q2 Sheet (for realism / complexity) ---
    ws2 = wb.create_sheet('Q2')
    q2_headers = ['ID', 'Department', 'Projected Budget']
    for col, h in enumerate(q2_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    q2_data = [
        [1, 'Engineering', 48000],
        [2, 'Marketing', 30500],
        [3, 'Human Resources', 16200],
        [4, 'Sales', 65000],
        [5, 'Finance', 9100],
    ]
    for r, row_data in enumerate(q2_data, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    ws2.column_dimensions['A'].width = 6
    ws2.column_dimensions['B'].width = 28
    ws2.column_dimensions['C'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
