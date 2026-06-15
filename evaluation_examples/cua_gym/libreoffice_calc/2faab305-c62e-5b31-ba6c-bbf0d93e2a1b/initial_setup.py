"""
Initial Setup: Create benefits cost comparison matrix with plan data
Task ID: calc_hr_058
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_058'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Plans ---
    ws1 = wb.active
    ws1.title = 'Plans'

    # Headers
    headers = ['Plan', 'Monthly Premium', 'Employer Share', 'Employee Share',
               'Deductible', 'Copay', 'OOP Max']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows
    data = [
        ['Gold', 850, 0.80, 0.20, 500, 20, 3000],
        ['Silver', 650, 0.75, 0.25, 1500, 35, 5000],
        ['Bronze', 450, 0.70, 0.30, 3000, 50, 7000],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Format percentage columns (C, D) as percentage
    for r in range(2, 5):
        ws1.cell(row=r, column=3).number_format = '0%'
        ws1.cell(row=r, column=4).number_format = '0%'

    # Format currency columns (B, E, F, G)
    for r in range(2, 5):
        ws1.cell(row=r, column=2).number_format = '$#,##0'
        ws1.cell(row=r, column=5).number_format = '$#,##0'
        ws1.cell(row=r, column=6).number_format = '$#,##0'
        ws1.cell(row=r, column=7).number_format = '$#,##0'

    # Adjust column widths
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 16
    ws1.column_dimensions['D'].width = 16
    ws1.column_dimensions['E'].width = 14
    ws1.column_dimensions['F'].width = 10
    ws1.column_dimensions['G'].width = 12

    # --- Sheet 2: CostCalc ---
    ws2 = wb.create_sheet('CostCalc')

    # Headers
    cost_headers = ['Plan', 'Annual Employee Cost', 'Annual Employer Cost']
    for col, h in enumerate(cost_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Plan names only - NO formulas (task is to add them)
    plan_names = ['Gold', 'Silver', 'Bronze']
    for r, name in enumerate(plan_names, 2):
        cell = ws2.cell(row=r, column=1, value=name)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Format columns B, C as currency (but leave empty - agent fills them)
    for r in range(2, 5):
        ws2.cell(row=r, column=2).number_format = '$#,##0.00'
        ws2.cell(row=r, column=3).number_format = '$#,##0.00'

    # Adjust column widths
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 24
    ws2.column_dimensions['C'].width = 24

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
