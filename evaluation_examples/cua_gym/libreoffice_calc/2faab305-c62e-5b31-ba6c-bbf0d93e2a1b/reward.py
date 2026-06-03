"""
Reward Script: Benefits cost comparison matrix with employee/employer contributions
Task ID: calc_hr_058
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.4): Employee annual cost formulas in CostCalc B2:B4
  - Component 2 (0.4): Employer annual cost formulas in CostCalc C2:C4
  - Component 3 (0.2): All 6 formulas present and complete
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_058'

# Ground truth values from task context:
# Gold employee annual: 850*0.2*12 + 500 = 2540, employer: 850*0.8*12 = 8160
# Silver employee: 650*0.25*12 + 1500 = 2450, employer: 650*0.75*12 = 5850
# Bronze employee: 450*0.3*12 + 3000 = 4620, employer: 450*0.7*12 = 3780

def normalize_formula(f):
    """Normalize a formula string for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ""
    return f.upper().replace(" ", "")

def is_employee_cost_formula(formula, row):
    """
    Check if formula computes: Premium * EmployeeShare * 12 + Deductible
    for the given row. Accepts various formula structures.
    Premium=Plans!B{row}, EmployeeShare=Plans!D{row}, Deductible=Plans!E{row}
    """
    f = normalize_formula(formula)
    if not f.startswith("="):
        return False

    # Must reference Premium (B), EmployeeShare (D), and Deductible (E) from Plans sheet
    has_premium = f"PLANS!B{row}" in f
    has_emp_share = f"PLANS!D{row}" in f
    has_deductible = f"PLANS!E{row}" in f
    has_12 = "12" in f

    return has_premium and has_emp_share and has_deductible and has_12

def is_employer_cost_formula(formula, row):
    """
    Check if formula computes: Premium * EmployerShare * 12
    for the given row.
    Premium=Plans!B{row}, EmployerShare=Plans!C{row}
    """
    f = normalize_formula(formula)
    if not f.startswith("="):
        return False

    has_premium = f"PLANS!B{row}" in f
    has_empr_share = f"PLANS!C{row}" in f
    has_12 = "12" in f

    return has_premium and has_empr_share and has_12

def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: CostCalc sheet must exist
    if 'CostCalc' not in wb.sheetnames:
        print("FAIL: 'CostCalc' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['CostCalc']

    # Component 1: Employee annual cost formulas in B2:B4 (0.4 points)
    # Formula: Premium * EmployeeShare * 12 + Deductible
    # Each correct formula = 0.4/3 points
    try:
        emp_pass = 0
        plan_rows = {2: 'Gold', 3: 'Silver', 4: 'Bronze'}
        for row, plan_name in plan_rows.items():
            formula = ws.cell(row=row, column=2).value
            if is_employee_cost_formula(formula, row):
                print(f"PASS: Component 1 — B{row} ({plan_name}) employee cost formula correct: {formula}")
                emp_pass += 1
            else:
                print(f"FAIL: Component 1 — B{row} ({plan_name}) expected employee cost formula, found: {repr(formula)}")

        if emp_pass > 0:
            points = round(0.4 * emp_pass / 3, 4)
            print(f"  Component 1 subtotal: {emp_pass}/3 formulas correct ({points} pts)")
            total_score += points
        else:
            print(f"  Component 1 subtotal: 0/3 formulas correct (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Employer annual cost formulas in C2:C4 (0.4 points)
    # Formula: Premium * EmployerShare * 12
    # Each correct formula = 0.4/3 points
    try:
        empr_pass = 0
        for row, plan_name in plan_rows.items():
            formula = ws.cell(row=row, column=3).value
            if is_employer_cost_formula(formula, row):
                print(f"PASS: Component 2 — C{row} ({plan_name}) employer cost formula correct: {formula}")
                empr_pass += 1
            else:
                print(f"FAIL: Component 2 — C{row} ({plan_name}) expected employer cost formula, found: {repr(formula)}")

        if empr_pass > 0:
            points = round(0.4 * empr_pass / 3, 4)
            print(f"  Component 2 subtotal: {empr_pass}/3 formulas correct ({points} pts)")
            total_score += points
        else:
            print(f"  Component 2 subtotal: 0/3 formulas correct (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Completeness — all 6 formula cells are non-empty formulas (0.2 points)
    # This checks that ALL cells have formulas (not just some), rewarding completeness
    try:
        all_filled = 0
        for row in [2, 3, 4]:
            for col in [2, 3]:
                val = ws.cell(row=row, column=col).value
                if isinstance(val, str) and val.startswith("="):
                    all_filled += 1

        if all_filled == 6:
            print(f"PASS: Component 3 — All 6 formula cells populated ({0.2} pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — Only {all_filled}/6 formula cells populated")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook for LibreOffice
def persist_app_state():
    import os, time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")

# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
