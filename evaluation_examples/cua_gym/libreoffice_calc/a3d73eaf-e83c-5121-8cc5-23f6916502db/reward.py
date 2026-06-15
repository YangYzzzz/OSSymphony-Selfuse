"""
Reward Script: SUMPRODUCT multi-criteria exact match lookup
Task ID: calc_lf_004
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): E2 contains a formula (starts with '=')
  Component 2 (0.3): The formula uses SUMPRODUCT function
  Component 3 (0.4): The formula correctly performs the multi-criteria lookup
                      (references "Sales", "West", and the Revenue column C2:C7)
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_004'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice document."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'DeptRevenue' sheet must exist
    if 'DeptRevenue' not in wb.sheetnames:
        print(f"CRITICAL: 'DeptRevenue' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['DeptRevenue']

    # Component 1: E2 contains a formula (0.3 points)
    # This FAILS on initial (E2 is empty) and PASSES on golden (E2 has formula)
    try:
        e2_value = ws['E2'].value
        if e2_value is not None and isinstance(e2_value, str) and e2_value.startswith('='):
            print(f"PASS: Component 1 — E2 contains a formula: {e2_value[:50]}... (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — E2 does not contain a formula. Value: {repr(e2_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: The formula uses SUMPRODUCT (0.3 points)
    # This FAILS on initial (no formula) and PASSES on golden (SUMPRODUCT formula)
    try:
        e2_value = ws['E2'].value
        if e2_value is not None and isinstance(e2_value, str):
            formula_upper = e2_value.upper().replace(" ", "")
            if 'SUMPRODUCT' in formula_upper:
                print(f"PASS: Component 2 — Formula uses SUMPRODUCT function (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — Formula does not use SUMPRODUCT. Formula: {e2_value}")
        else:
            print(f"FAIL: Component 2 — E2 is not a formula string. Value: {repr(e2_value)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Formula correctly references "Sales", "West", and C column range (0.4 points)
    # The formula must produce 58000 for dept=Sales, region=West.
    # We verify the formula contains the right criteria and range references.
    try:
        e2_value = ws['E2'].value
        if e2_value is not None and isinstance(e2_value, str):
            formula_normalized = e2_value.upper().replace(" ", "")

            sub_score = 0.0
            checks_passed = 0

            # Check 3a: References "Sales" criterion
            if '"SALES"' in formula_normalized or "'SALES'" in formula_normalized:
                checks_passed += 1

            # Check 3b: References "West" criterion
            if '"WEST"' in formula_normalized or "'WEST'" in formula_normalized:
                checks_passed += 1

            # Check 3c: References Revenue column range (C2:C7 or similar C range)
            if re.search(r'C\d+:C\d+', formula_normalized):
                checks_passed += 1

            # Check 3d: References Department column range (A2:A7 or similar A range)
            if re.search(r'A\d+:A\d+', formula_normalized):
                checks_passed += 1

            if checks_passed == 4:
                print(f"PASS: Component 3 — Formula has correct criteria and ranges (0.4 pts)")
                total_score += 0.4
            elif checks_passed >= 2:
                partial = round(0.4 * checks_passed / 4, 2)
                print(f"PARTIAL: Component 3 — {checks_passed}/4 sub-checks passed ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 3 — Only {checks_passed}/4 sub-checks passed. Formula: {e2_value}")
        else:
            print(f"FAIL: Component 3 — E2 is not a formula string. Value: {repr(e2_value)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist any unsaved GUI edits before verification
persist_app_state("libreoffice_calc")

# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
