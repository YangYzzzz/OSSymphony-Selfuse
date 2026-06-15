"""
Initial Setup: SUMPRODUCT multi-criteria lookup task
Task ID: calc_lf_004
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_004'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: DeptRevenue ---
    ws = wb.active
    ws.title = 'DeptRevenue'

    # Headers
    ws['A1'] = 'Department'
    ws['B1'] = 'Region'
    ws['C1'] = 'Revenue'

    # Data rows (exactly as specified in context)
    data = [
        ['Sales',     'East',  45000],
        ['Marketing', 'West',  32000],
        ['Sales',     'West',  58000],
        ['Sales',     'North', 41000],
        ['Marketing', 'East',  29000],
        ['Sales',     'East',  47000],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Result label in E1, E2 left empty (task asks agent to enter formula there)
    ws['E1'] = 'Result'
    # E2 is intentionally empty

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['E'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
