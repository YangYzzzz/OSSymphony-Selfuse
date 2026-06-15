"""
Initial Setup: Insert a new column A with Row ID header and sequential numbers
Task ID: calc_cop_insert_row_col_007
Domain: libreoffice_calc

Creates a spreadsheet with a single sheet 'Contacts' containing:
- Column A: First Name (header in A1)
- Column B: Last Name
- Column C: Email
- 50 data rows (rows 2-51)
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_cop_insert_row_col_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Contacts ---
    ws = wb.active
    ws.title = 'Contacts'

    # Headers in row 1
    ws['A1'] = 'First Name'
    ws['B1'] = 'Last Name'
    ws['C1'] = 'Email'

    # Realistic contact data — 50 rows (rows 2-51)
    contacts = [
        ('Sarah', 'Chen', 'sarah.chen@techcorp.com'),
        ('Marcus', 'Johnson', 'marcus.johnson@mediagroup.net'),
        ('Emily', 'Rodriguez', 'e.rodriguez@financeplus.org'),
        ('David', 'Kim', 'david.kim@innovate.io'),
        ('Priya', 'Patel', 'priya.patel@globalhealth.org'),
        ('James', 'Williams', 'j.williams@retailco.com'),
        ('Aisha', 'Thompson', 'aisha.t@lawfirm.net'),
        ('Robert', 'Nakamura', 'rnakamura@architecturelab.com'),
        ('Sophia', 'Martinez', 'sophia.m@educationhub.edu'),
        ('Liam', 'Anderson', 'l.anderson@cloudtech.io'),
        ('Fatima', 'Ali', 'fatima.ali@medcenter.org'),
        ('Noah', 'Brown', 'noah.brown@logistics.net'),
        ('Zoe', 'Davis', 'zoe.davis@designstudio.com'),
        ('Ethan', 'Wilson', 'ethan.wilson@energyco.com'),
        ('Mia', 'Garcia', 'mia.garcia@consulting.biz'),
        ('Oliver', 'Lee', 'oliver.lee@banking.com'),
        ('Isabella', 'Taylor', 'i.taylor@pharma.com'),
        ('William', 'Moore', 'w.moore@agritech.net'),
        ('Chloe', 'Jackson', 'chloe.jackson@telecom.com'),
        ('Benjamin', 'White', 'ben.white@realestate.org'),
        ('Amara', 'Harris', 'amara.harris@nonprofit.org'),
        ('Henry', 'Clark', 'henry.clark@automotive.com'),
        ('Elena', 'Lewis', 'elena.lewis@hospitality.net'),
        ('Sebastian', 'Robinson', 's.robinson@security.com'),
        ('Natalia', 'Walker', 'natalia.walker@biotech.com'),
        ('Alexander', 'Hall', 'alex.hall@aerospace.net'),
        ('Leila', 'Young', 'leila.young@fashion.com'),
        ('Lucas', 'King', 'lucas.king@publishing.net'),
        ('Aria', 'Wright', 'aria.wright@eventmgmt.com'),
        ('Mason', 'Scott', 'mason.scott@insurance.com'),
        ('Harper', 'Green', 'harper.green@foodtech.net'),
        ('Elijah', 'Adams', 'elijah.adams@sportsmgmt.com'),
        ('Avery', 'Baker', 'avery.baker@construction.org'),
        ('Jackson', 'Gonzalez', 'j.gonzalez@shipping.com'),
        ('Luna', 'Nelson', 'luna.nelson@softwarelab.io'),
        ('Aiden', 'Carter', 'aiden.carter@healthtech.com'),
        ('Grace', 'Mitchell', 'grace.mitchell@lawassoc.net'),
        ('Caleb', 'Perez', 'caleb.perez@digitalagency.com'),
        ('Stella', 'Roberts', 'stella.roberts@dataanalytics.net'),
        ('Owen', 'Turner', 'owen.turner@envirotech.org'),
        ('Violet', 'Phillips', 'violet.phillips@hrservices.com'),
        ('Isaiah', 'Campbell', 'isaiah.campbell@venturecap.com'),
        ('Aurora', 'Parker', 'aurora.parker@creativemedia.com'),
        ('Lincoln', 'Evans', 'lincoln.evans@urbandev.net'),
        ('Hazel', 'Edwards', 'hazel.edwards@cybersecurity.com'),
        ('Nolan', 'Collins', 'nolan.collins@airtravel.net'),
        ('Layla', 'Stewart', 'layla.stewart@publicrelations.com'),
        ('Carter', 'Sanchez', 'carter.sanchez@blockchain.io'),
        ('Riley', 'Morris', 'riley.morris@sustainability.org'),
        ('Wyatt', 'Rogers', 'wyatt.rogers@marketresearch.com'),
    ]

    for row_idx, (first, last, email) in enumerate(contacts, 2):
        ws.cell(row=row_idx, column=1, value=first)
        ws.cell(row=row_idx, column=2, value=last)
        ws.cell(row=row_idx, column=3, value=email)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Contacts')
    print(f'  Headers: First Name (A1), Last Name (B1), Email (C1)')
    print(f'  Data rows: {len(contacts)} (rows 2-51)')


create_initial()
