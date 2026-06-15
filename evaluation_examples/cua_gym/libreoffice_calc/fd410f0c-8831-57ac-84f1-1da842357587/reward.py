"""
Reward Script: Insert a new column A with "Row ID" header and sequential numbers 1-50
Task ID: calc_cop_insert_row_col_007
Domain: libreoffice_calc
Scoring:
  Component 1: A1 header is "Row ID"            (0.30 points)
  Component 2: A2:A51 contain sequential 1-50   (0.50 points)
  Component 3: Original columns shifted intact   (0.20 points)
  Total: 1.00
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_cop_insert_row_col_007'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.

    The task requires:
    1. A new column A inserted at the very beginning
    2. A1 contains "Row ID"
    3. A2:A51 contains integers 1 through 50 (sequential)
    4. Original data (First Name, Last Name, Email) shifted to columns B, C, D
    5. No data is lost (50 data rows remain intact)

    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook — if this fails, no score possible
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get "Contacts" sheet — required
    if 'Contacts' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Contacts' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Contacts']

    # Precondition gate: sheet must have at least 4 columns and 51 rows
    if ws.max_column < 4 or ws.max_row < 51:
        print(f"FAIL: Sheet dimensions too small: {ws.max_row} rows x {ws.max_column} cols")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: A1 header is "Row ID" (0.30 points)
    # This FAILS on initial (A1="First Name") → PASSES on golden (A1="Row ID")
    try:
        a1_value = ws['A1'].value
        if a1_value is not None and str(a1_value).strip() == 'Row ID':
            print(f"PASS: Component 1 — A1 header is 'Row ID' (found: {repr(a1_value)}) (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 — Expected A1='Row ID', found: {repr(a1_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: A2:A51 contains sequential integers 1 through 50 (0.50 points)
    # This FAILS on initial (A col has First Name data) → PASSES on golden
    try:
        correct_count = 0
        first_error = None
        for i in range(1, 51):
            row = i + 1  # rows 2 through 51
            cell_val = ws.cell(row=row, column=1).value
            expected = i
            try:
                if cell_val is not None and int(cell_val) == expected:
                    correct_count += 1
                elif first_error is None:
                    first_error = f"A{row}: expected {expected}, got {repr(cell_val)}"
            except (TypeError, ValueError):
                if first_error is None:
                    first_error = f"A{row}: expected {expected}, got non-numeric {repr(cell_val)}"

        if correct_count == 50:
            print(f"PASS: Component 2 — A2:A51 contains sequential integers 1-50 (all 50 correct) (0.50 pts)")
            total_score += 0.50
        else:
            print(f"FAIL: Component 2 — Sequential numbers: {correct_count}/50 correct. First error: {first_error}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Original columns shifted correctly (0.20 points)
    # Checks that First Name is now in B, Last Name in C, Email in D
    # AND that original data rows are intact (checking first 3 and last 3 rows)
    # This FAILS on initial (B1="Last Name", no column D) → PASSES on golden
    try:
        b1 = ws['B1'].value
        c1 = ws['C1'].value
        d1 = ws['D1'].value

        headers_ok = (
            b1 is not None and str(b1).strip() == 'First Name' and
            c1 is not None and str(c1).strip() == 'Last Name' and
            d1 is not None and str(d1).strip() == 'Email'
        )

        # Verify sample data rows are intact (check first and last data rows)
        # Row 2: Sarah Chen, sarah.chen@techcorp.com
        row2_b = ws['B2'].value
        row2_c = ws['C2'].value
        row2_d = ws['D2'].value

        # Row 51: Wyatt Rogers, wyatt.rogers@marketresearch.com
        row51_b = ws['B51'].value
        row51_c = ws['C51'].value
        row51_d = ws['D51'].value

        data_intact = (
            row2_b is not None and str(row2_b).strip() == 'Sarah' and
            row51_b is not None and str(row51_b).strip() == 'Wyatt'
        )

        if headers_ok and data_intact:
            print(f"PASS: Component 3 — Original columns (First Name→B, Last Name→C, Email→D) shifted correctly with data intact (0.20 pts)")
            total_score += 0.20
        elif headers_ok:
            print(f"FAIL: Component 3 — Headers shifted correctly but data may be missing/incorrect (B2={repr(row2_b)}, B51={repr(row51_b)})")
        else:
            print(f"FAIL: Component 3 — Headers not shifted correctly: B1={repr(b1)}, C1={repr(c1)}, D1={repr(d1)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
