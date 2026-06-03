"""
Initial Setup: Financial model with manual calculation mode enabled.
Task ID: calc_tbl_015
Domain: libreoffice_calc

Creates a financial model spreadsheet with 200+ formulas.
Row 50 contains summary totals via SUM formulas.
Sets calcMode="manual" directly in the xlsx workbook.xml so LibreOffice opens
with automatic recalculation disabled.
"""

import os
import re
import shlex
import subprocess
import time
import glob
import zipfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_015'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def set_calc_mode_manual(filepath):
    """
    Set calcMode='manual' directly in the xlsx workbook.xml.
    This is the OOXML-level way to disable automatic calculation.
    """
    temp_path = filepath + '.tmp'
    with zipfile.ZipFile(filepath, 'r') as zin:
        with zipfile.ZipFile(temp_path, 'w') as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == 'xl/workbook.xml':
                    content = data.decode('utf-8')
                    if 'calcMode=' in content:
                        content = re.sub(r'calcMode="[^"]*"', 'calcMode="manual"', content)
                    elif 'calcPr' in content:
                        content = content.replace('<calcPr ', '<calcPr calcMode="manual" ')
                    else:
                        content = content.replace('</workbook>', '<calcPr calcMode="manual"/></workbook>')
                    data = content.encode('utf-8')
                zout.writestr(item, data)
    os.replace(temp_path, filepath)
    print(f'Set calcMode="manual" in {filepath}')


def create_initial():
    wb = openpyxl.Workbook()

    # --- Style definitions ---
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    currency_fmt = '$#,##0.00'
    pct_fmt = '0.0%'
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    total_font = Font(name="Calibri", size=11, bold=True)
    total_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")

    # =============================================
    # Sheet 1: Revenue Projections
    # =============================================
    ws1 = wb.active
    ws1.title = "Revenue Projections"

    # Headers
    rev_headers = ["Product Line", "Q1 Revenue", "Q2 Revenue", "Q3 Revenue", "Q4 Revenue",
                   "Annual Total", "YoY Growth", "Margin %", "Gross Profit"]
    for c, h in enumerate(rev_headers, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    ws1.column_dimensions['A'].width = 24
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws1.column_dimensions[col_letter].width = 15

    # Product data (rows 2-49) -- 48 product lines
    products = [
        ("Enterprise SaaS Platform", 245000, 268000, 289000, 312000, 0.15, 0.72),
        ("Cloud Infrastructure", 189000, 201000, 218000, 235000, 0.12, 0.65),
        ("Data Analytics Suite", 156000, 172000, 185000, 198000, 0.18, 0.78),
        ("Cybersecurity Services", 134000, 145000, 158000, 172000, 0.22, 0.68),
        ("Mobile Dev Toolkit", 98000, 105000, 112000, 121000, 0.10, 0.74),
        ("AI/ML Platform", 210000, 238000, 265000, 298000, 0.28, 0.70),
        ("DevOps Automation", 87000, 94000, 102000, 110000, 0.14, 0.66),
        ("IoT Management Hub", 76000, 82000, 89000, 96000, 0.11, 0.62),
        ("API Gateway Pro", 112000, 121000, 130000, 140000, 0.16, 0.75),
        ("Identity Management", 95000, 103000, 111000, 119000, 0.13, 0.71),
        ("Database Solutions", 145000, 156000, 168000, 181000, 0.17, 0.69),
        ("Edge Computing Suite", 68000, 76000, 85000, 95000, 0.25, 0.63),
        ("Compliance Manager", 54000, 58000, 63000, 68000, 0.09, 0.80),
        ("Video Conferencing", 178000, 190000, 204000, 219000, 0.14, 0.58),
        ("Project Management", 132000, 142000, 153000, 165000, 0.15, 0.72),
        ("CRM Integration", 108000, 116000, 125000, 135000, 0.13, 0.67),
        ("HR Tech Platform", 89000, 96000, 104000, 112000, 0.16, 0.73),
        ("Supply Chain Analytics", 125000, 135000, 146000, 157000, 0.19, 0.64),
        ("Digital Marketing Tools", 92000, 100000, 108000, 117000, 0.17, 0.76),
        ("E-Commerce Engine", 156000, 168000, 181000, 195000, 0.15, 0.61),
        ("Payment Processing", 198000, 213000, 229000, 246000, 0.14, 0.55),
        ("Warehouse Management", 73000, 79000, 85000, 92000, 0.11, 0.68),
        ("Customer Support AI", 115000, 126000, 138000, 151000, 0.21, 0.77),
        ("Network Monitoring", 67000, 72000, 78000, 84000, 0.10, 0.70),
        ("Content Delivery", 142000, 153000, 165000, 177000, 0.13, 0.59),
        ("Blockchain Services", 48000, 55000, 63000, 72000, 0.32, 0.82),
        ("AR/VR Development", 62000, 71000, 81000, 93000, 0.30, 0.65),
        ("Quantum Computing API", 35000, 42000, 50000, 59000, 0.35, 0.85),
        ("Green IT Solutions", 58000, 64000, 71000, 79000, 0.20, 0.74),
        ("Healthcare Data", 167000, 180000, 194000, 209000, 0.16, 0.71),
        ("FinTech Middleware", 143000, 154000, 166000, 179000, 0.15, 0.66),
        ("Robotic Process Auto", 88000, 97000, 107000, 118000, 0.23, 0.73),
        ("Smart Building IoT", 52000, 57000, 63000, 69000, 0.18, 0.60),
        ("AgriTech Analytics", 41000, 45000, 50000, 55000, 0.14, 0.69),
        ("EdTech Platform", 78000, 85000, 92000, 100000, 0.17, 0.75),
        ("Travel Tech Suite", 94000, 102000, 110000, 119000, 0.13, 0.62),
        ("Insurance Automation", 110000, 119000, 128000, 138000, 0.16, 0.70),
        ("Real Estate Portal", 86000, 93000, 100000, 108000, 0.15, 0.64),
        ("Gaming Backend", 72000, 80000, 89000, 99000, 0.24, 0.57),
        ("Sustainability Track", 45000, 50000, 56000, 62000, 0.21, 0.78),
        ("Legal Tech Platform", 63000, 68000, 74000, 80000, 0.12, 0.81),
        ("Logistics Optimizer", 98000, 106000, 115000, 124000, 0.14, 0.63),
        ("Food Safety Monitor", 37000, 40000, 44000, 48000, 0.11, 0.72),
        ("Sports Analytics", 55000, 60000, 66000, 72000, 0.19, 0.67),
        ("Telecom Management", 128000, 138000, 149000, 160000, 0.13, 0.60),
        ("Energy Grid Optimize", 83000, 90000, 98000, 106000, 0.17, 0.69),
        ("Retail POS System", 105000, 113000, 122000, 131000, 0.12, 0.65),
        ("Media Streaming API", 91000, 99000, 108000, 117000, 0.18, 0.58),
    ]

    for r, (name, q1, q2, q3, q4, yoy, margin) in enumerate(products, 2):
        ws1.cell(row=r, column=1, value=name).border = thin_border
        ws1.cell(row=r, column=2, value=q1).number_format = currency_fmt
        ws1.cell(row=r, column=2).border = thin_border
        ws1.cell(row=r, column=3, value=q2).number_format = currency_fmt
        ws1.cell(row=r, column=3).border = thin_border
        ws1.cell(row=r, column=4, value=q3).number_format = currency_fmt
        ws1.cell(row=r, column=4).border = thin_border
        ws1.cell(row=r, column=5, value=q4).number_format = currency_fmt
        ws1.cell(row=r, column=5).border = thin_border
        # Annual Total = sum of Q1-Q4 (formula)
        ws1.cell(row=r, column=6, value=f'=SUM(B{r}:E{r})').number_format = currency_fmt
        ws1.cell(row=r, column=6).border = thin_border
        # YoY Growth
        ws1.cell(row=r, column=7, value=yoy).number_format = pct_fmt
        ws1.cell(row=r, column=7).border = thin_border
        # Margin %
        ws1.cell(row=r, column=8, value=margin).number_format = pct_fmt
        ws1.cell(row=r, column=8).border = thin_border
        # Gross Profit = Annual Total * Margin (formula)
        ws1.cell(row=r, column=9, value=f'=F{r}*H{r}').number_format = currency_fmt
        ws1.cell(row=r, column=9).border = thin_border

    # Row 50: Summary totals with formulas
    total_row = 50
    ws1.cell(row=total_row, column=1, value="TOTAL").font = total_font
    ws1.cell(row=total_row, column=1).fill = total_fill
    ws1.cell(row=total_row, column=1).border = thin_border

    for col in range(2, 6):  # B-E: SUM of quarterly revenues
        col_letter = openpyxl.utils.get_column_letter(col)
        ws1.cell(row=total_row, column=col,
                 value=f'=SUM({col_letter}2:{col_letter}49)')
        ws1.cell(row=total_row, column=col).number_format = currency_fmt
        ws1.cell(row=total_row, column=col).font = total_font
        ws1.cell(row=total_row, column=col).fill = total_fill
        ws1.cell(row=total_row, column=col).border = thin_border

    # F50: Annual Total sum
    ws1.cell(row=total_row, column=6, value='=SUM(F2:F49)')
    ws1.cell(row=total_row, column=6).number_format = currency_fmt
    ws1.cell(row=total_row, column=6).font = total_font
    ws1.cell(row=total_row, column=6).fill = total_fill
    ws1.cell(row=total_row, column=6).border = thin_border

    # G50: Average YoY Growth
    ws1.cell(row=total_row, column=7, value='=AVERAGE(G2:G49)')
    ws1.cell(row=total_row, column=7).number_format = pct_fmt
    ws1.cell(row=total_row, column=7).font = total_font
    ws1.cell(row=total_row, column=7).fill = total_fill
    ws1.cell(row=total_row, column=7).border = thin_border

    # H50: Average Margin
    ws1.cell(row=total_row, column=8, value='=AVERAGE(H2:H49)')
    ws1.cell(row=total_row, column=8).number_format = pct_fmt
    ws1.cell(row=total_row, column=8).font = total_font
    ws1.cell(row=total_row, column=8).fill = total_fill
    ws1.cell(row=total_row, column=8).border = thin_border

    # I50: Total Gross Profit
    ws1.cell(row=total_row, column=9, value='=SUM(I2:I49)')
    ws1.cell(row=total_row, column=9).number_format = currency_fmt
    ws1.cell(row=total_row, column=9).font = total_font
    ws1.cell(row=total_row, column=9).fill = total_fill
    ws1.cell(row=total_row, column=9).border = thin_border

    # Freeze header row
    ws1.freeze_panes = "A2"

    # =============================================
    # Sheet 2: Expense Breakdown
    # =============================================
    ws2 = wb.create_sheet("Expense Breakdown")

    exp_headers = ["Department", "Salaries", "Equipment", "Software", "Travel",
                   "Training", "Misc", "Dept Total"]
    for c, h in enumerate(exp_headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws2.column_dimensions['A'].width = 22
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws2.column_dimensions[col_letter].width = 14

    departments = [
        ("Engineering", 1850000, 320000, 185000, 45000, 28000, 12000),
        ("Sales & Marketing", 1240000, 95000, 125000, 180000, 35000, 22000),
        ("Product Management", 980000, 45000, 110000, 62000, 25000, 8000),
        ("Customer Success", 720000, 35000, 85000, 38000, 22000, 6000),
        ("Human Resources", 540000, 25000, 65000, 18000, 42000, 9000),
        ("Finance & Legal", 680000, 30000, 95000, 25000, 18000, 15000),
        ("Operations", 460000, 180000, 72000, 32000, 15000, 25000),
        ("Research & Dev", 1100000, 450000, 210000, 55000, 48000, 18000),
        ("IT Infrastructure", 890000, 520000, 290000, 28000, 32000, 14000),
        ("Executive Office", 1200000, 15000, 35000, 95000, 12000, 35000),
    ]

    for r, (dept, *vals) in enumerate(departments, 2):
        ws2.cell(row=r, column=1, value=dept).border = thin_border
        for c, v in enumerate(vals, 2):
            ws2.cell(row=r, column=c, value=v).number_format = currency_fmt
            ws2.cell(row=r, column=c).border = thin_border
        # Dept Total formula
        ws2.cell(row=r, column=8, value=f'=SUM(B{r}:G{r})').number_format = currency_fmt
        ws2.cell(row=r, column=8).border = thin_border

    # Summary row for expenses
    exp_total_row = 12
    ws2.cell(row=exp_total_row, column=1, value="TOTAL").font = total_font
    ws2.cell(row=exp_total_row, column=1).fill = total_fill
    ws2.cell(row=exp_total_row, column=1).border = thin_border
    for col in range(2, 9):
        letter = openpyxl.utils.get_column_letter(col)
        ws2.cell(row=exp_total_row, column=col, value=f'=SUM({letter}2:{letter}11)')
        ws2.cell(row=exp_total_row, column=col).number_format = currency_fmt
        ws2.cell(row=exp_total_row, column=col).font = total_font
        ws2.cell(row=exp_total_row, column=col).fill = total_fill
        ws2.cell(row=exp_total_row, column=col).border = thin_border

    # =============================================
    # Sheet 3: KPI Dashboard
    # =============================================
    ws3 = wb.create_sheet("KPI Dashboard")

    kpi_headers = ["Metric", "Target", "Actual", "Variance", "Status"]
    for c, h in enumerate(kpi_headers, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws3.column_dimensions['A'].width = 30
    for col_letter in ['B', 'C', 'D', 'E']:
        ws3.column_dimensions[col_letter].width = 16

    kpis = [
        ("Monthly Recurring Revenue", 850000, 892000),
        ("Customer Acquisition Cost", 1200, 1150),
        ("Customer Lifetime Value", 48000, 52000),
        ("Churn Rate", 0.025, 0.021),
        ("Net Promoter Score", 72, 78),
        ("Employee Satisfaction", 4.2, 4.5),
        ("Gross Margin", 0.68, 0.71),
        ("Operating Margin", 0.22, 0.25),
        ("Revenue per Employee", 285000, 298000),
        ("Customer Satisfaction", 0.92, 0.94),
    ]

    for r, (metric, target, actual) in enumerate(kpis, 2):
        ws3.cell(row=r, column=1, value=metric).border = thin_border
        ws3.cell(row=r, column=2, value=target).border = thin_border
        ws3.cell(row=r, column=3, value=actual).border = thin_border
        # Variance formula
        ws3.cell(row=r, column=4, value=f'=C{r}-B{r}').border = thin_border
        # Status formula
        ws3.cell(row=r, column=5, value=f'=IF(C{r}>=B{r},"On Track","Below Target")').border = thin_border

    # Format numeric cells
    for r in range(2, 12):
        for c in [2, 3, 4]:
            ws3.cell(row=r, column=c).number_format = '#,##0.00'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Set calcMode="manual" at the OOXML level
    set_calc_mode_manual(OUTPUT)

    # Verify the setting
    with zipfile.ZipFile(OUTPUT, 'r') as z:
        with z.open('xl/workbook.xml') as f:
            content = f.read().decode('utf-8')
            if 'calcMode="manual"' in content:
                print('VERIFIED: calcMode="manual" in workbook.xml')
            else:
                print('WARNING: calcMode not set correctly')

    # GUI-ready startup: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
