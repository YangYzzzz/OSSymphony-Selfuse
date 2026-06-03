"""
Reward Script: Switch calculation mode from manual to automatic
Task ID: calc_tbl_015
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): calcMode is NOT 'manual' (automatic calculation enabled)
  Component 2 (0.3): Row 50 cached values are NOT stale zeros (recalculated)
  Component 3 (0.2): Calc auto AND row 50 formulas intact (compound check)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_015'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook for formula inspection
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Load workbook again with data_only=True for cached values
    try:
        wb_data = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"CRITICAL: Cannot load file with data_only=True: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active
    ws_data = wb_data.active

    # Precondition: file has the expected structure (row 50 with TOTAL label)
    if ws.cell(50, 1).value != "TOTAL":
        print("FAIL: Precondition — Row 50 column 1 should be 'TOTAL', "
              f"found: {ws.cell(50, 1).value}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Calculation mode is NOT 'manual' (0.5 points)
    # In the initial state, calcMode='manual'. The task requires switching to automatic.
    # In openpyxl, automatic mode is represented as calcMode=None (the default).
    try:
        calc_props = wb.calculation
        if calc_props is not None:
            calc_mode = calc_props.calcMode
        else:
            calc_mode = None

        if calc_mode != 'manual':
            print(f"PASS: Component 1 — calcMode is '{calc_mode}' (not 'manual') (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — calcMode is still 'manual', expected automatic (None)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Row 50 cached values are NOT stale zeros (0.3 points)
    # In the initial state, cached values for row 50 are all 0 (stale).
    # After switching to automatic and recalculating, cached values should change.
    # With openpyxl, they may be None (if not cached) or actual computed values,
    # but they should NOT be 0 for all columns.
    try:
        stale_count = 0
        checked_cols = 0
        for c in range(2, 10):  # columns B through I (2-9)
            cached_val = ws_data.cell(50, c).value
            if cached_val == 0:
                stale_count += 1
            checked_cols += 1

        # In the initial state, ALL 8 columns have cached value 0.
        # In the golden state, they should NOT all be 0.
        # We check that fewer than all columns have stale 0 values.
        if stale_count < checked_cols:
            print(f"PASS: Component 2 — Row 50 cached values are not all stale zeros "
                  f"({stale_count}/{checked_cols} are 0) (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — All {checked_cols} cached values in row 50 are 0 (stale)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Calc mode is automatic AND row 50 formulas are still intact (0.2 points)
    # Compound check: only awards points when the primary task change (calcMode != manual)
    # is present AND the formulas haven't been corrupted. This ensures this component
    # FAILS on initial_env (where calcMode is still manual) and PASSES on golden_env.
    try:
        # First gate: calcMode must not be manual (anchored to task change)
        cp = wb.calculation
        is_auto = (cp is None) or (cp.calcMode != 'manual')

        if not is_auto:
            print("FAIL: Component 3 — calcMode is still 'manual', compound check fails")
        else:
            expected_formulas = {
                2: '=SUM(B2:B49)',
                3: '=SUM(C2:C49)',
                4: '=SUM(D2:D49)',
                5: '=SUM(E2:E49)',
                6: '=SUM(F2:F49)',
                7: '=AVERAGE(G2:G49)',
                8: '=AVERAGE(H2:H49)',
                9: '=SUM(I2:I49)',
            }

            formula_ok_count = 0
            total_formulas = len(expected_formulas)

            for col, expected in expected_formulas.items():
                actual = ws.cell(50, col).value
                if isinstance(actual, str) and actual.upper().replace(' ', '') == expected.upper().replace(' ', ''):
                    formula_ok_count += 1
                else:
                    print(f"  INFO: Col {col} formula mismatch — expected '{expected}', found '{actual}'")

            if formula_ok_count == total_formulas:
                print(f"PASS: Component 3 — Calc auto + all {total_formulas} formulas intact (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 3 — Only {formula_ok_count}/{total_formulas} formulas intact")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in a given env
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
