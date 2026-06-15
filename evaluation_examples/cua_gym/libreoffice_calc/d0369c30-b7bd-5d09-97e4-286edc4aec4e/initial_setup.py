"""
Initial Setup: Define a named range 'SalesData' covering B2:B101 on the Sales sheet
Task ID: calc_gg1_009
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_009'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    # Header in B1
    ws.cell(row=1, column=2, value="Monthly Sales")
    # Also add a label column header in A1
    ws.cell(row=1, column=1, value="Month")

    # Generate 100 rows of realistic monthly sales data (B2:B101)
    # Use a mix of realistic sales figures spanning about 8 years of monthly data
    random.seed(42)
    base_sales = 12500.0
    months = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]

    for i in range(100):
        row = i + 2
        year = 2016 + (i // 12)
        month_name = months[i % 12]
        ws.cell(row=row, column=1, value=f"{month_name} {year}")

        # Realistic sales with seasonal variation and growth trend
        seasonal_factor = 1.0 + 0.15 * (1 if i % 12 in [10, 11, 0] else 0)  # Q4/Jan bump
        growth = 1.0 + 0.02 * (i // 12)  # 2% annual growth
        noise = random.uniform(0.85, 1.15)
        sales = round(base_sales * seasonal_factor * growth * noise, 2)
        ws.cell(row=row, column=2, value=sales)

    # Set reasonable column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 15

    # NO named ranges defined - that is the task for the agent

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
