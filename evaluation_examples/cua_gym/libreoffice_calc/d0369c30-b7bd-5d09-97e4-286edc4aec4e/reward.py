"""
Reward Script: Define a named range 'SalesData' covering B2:B101 on the 'Sales' sheet
Task ID: calc_gg1_009
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): Named range 'SalesData' exists in the workbook
  Component 2 (0.3): Named range references exactly Sales!$B$2:$B$101
  Component 3 (0.2): Named range exists AND underlying data is intact (B1 header + B2:B101 populated)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_009'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Sales' sheet must exist (this is a precondition, not scored)
    if 'Sales' not in wb.sheetnames:
        print("CRITICAL: 'Sales' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Named range 'SalesData' exists (0.5 points)
    # This is the core task requirement - defining the named range
    try:
        has_salesdata = 'SalesData' in wb.defined_names
        if has_salesdata:
            print(f"PASS: Component 1 — Named range 'SalesData' exists in workbook (0.5 pts)")
            total_score += 0.5
        else:
            # Also check case-insensitive since user might type differently
            found_name = None
            for name in wb.defined_names:
                if name.lower() == 'salesdata':
                    found_name = name
                    break
            if found_name:
                print(f"PARTIAL: Component 1 — Found named range '{found_name}' (case mismatch, expected 'SalesData') (0.3 pts)")
                total_score += 0.3
            else:
                all_names = list(wb.defined_names.keys()) if hasattr(wb.defined_names, 'keys') else list(wb.defined_names)
                print(f"FAIL: Component 1 — Named range 'SalesData' not found. Existing names: {all_names}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Named range references exactly Sales!$B$2:$B$101 (0.3 points)
    # Verifies the range is correctly defined to cover the sales data
    try:
        if 'SalesData' in wb.defined_names:
            dn = wb.defined_names['SalesData']
            ref = dn.attr_text
            # Normalize: remove quotes around sheet name if present, compare case-insensitively
            normalized_ref = ref.replace("'", "").replace(" ", "").upper()
            expected_ref = "SALES!$B$2:$B$101"
            if normalized_ref == expected_ref:
                print(f"PASS: Component 2 — SalesData references '{ref}' which matches expected Sales!$B$2:$B$101 (0.3 pts)")
                total_score += 0.3
            else:
                # Check for common close variants (e.g., relative references)
                # Accept Sales!B2:B101 without $ signs as partial credit
                stripped = normalized_ref.replace("$", "")
                expected_stripped = expected_ref.replace("$", "")
                if stripped == expected_stripped:
                    print(f"PARTIAL: Component 2 — SalesData references '{ref}' (relative refs, expected absolute) (0.15 pts)")
                    total_score += 0.15
                else:
                    print(f"FAIL: Component 2 — SalesData references '{ref}', expected 'Sales!$B$2:$B$101'")
        else:
            # Check case-insensitive match
            found_dn = None
            for name, dn_obj in wb.defined_names.items():
                if name.lower() == 'salesdata':
                    found_dn = dn_obj
                    break
            if found_dn:
                ref = found_dn.attr_text
                normalized_ref = ref.replace("'", "").replace(" ", "").upper()
                expected_ref = "SALES!$B$2:$B$101"
                if normalized_ref == expected_ref:
                    print(f"PASS: Component 2 — Named range references correct range '{ref}' (0.3 pts)")
                    total_score += 0.3
                else:
                    print(f"FAIL: Component 2 — Named range references '{ref}', expected 'Sales!$B$2:$B$101'")
            else:
                print(f"FAIL: Component 2 — No 'SalesData' named range found to check reference")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Named range exists AND data integrity is maintained (0.2 points)
    # This is a compound check: named range must exist, AND the underlying data must be intact
    # The data integrity alone is a precondition, but combined with named range it verifies
    # the task was done without breaking existing data
    try:
        has_name = ('SalesData' in wb.defined_names) or any(
            n.lower() == 'salesdata' for n in wb.defined_names
        )

        if has_name:
            ws = wb['Sales']
            # Check header
            header_ok = ws['B1'].value == 'Monthly Sales'
            # Check data exists in B2 and B101 (first and last data rows)
            b2_val = ws['B2'].value
            b101_val = ws['B101'].value
            data_present = (b2_val is not None and b101_val is not None)
            # Check that data cells are numeric
            data_numeric = False
            if data_present:
                try:
                    data_numeric = isinstance(float(b2_val), float) and isinstance(float(b101_val), float)
                except (ValueError, TypeError):
                    data_numeric = False

            if header_ok and data_present and data_numeric:
                print(f"PASS: Component 3 — Named range exists AND data intact: header='Monthly Sales', B2={b2_val}, B101={b101_val} (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 3 — Data integrity issue: header_ok={header_ok}, data_present={data_present}, data_numeric={data_numeric}")
        else:
            print(f"FAIL: Component 3 — No named range found, compound check fails")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
