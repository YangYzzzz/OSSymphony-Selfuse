"""
Reward Script: Write IF formulas in column C to check if due dates have passed
Task ID: calc_fma_date_logic_048
Domain: libreoffice_calc
Scoring:
  Component 1: At least one IF formula exists in C2:C12 referencing TODAY()  (0.3 pts)
  Component 2: ALL 11 cells C2:C12 contain valid IF formulas with TODAY()     (0.4 pts)
  Component 3: Every formula uses correct structure =IF(Bx<TODAY(), overdue/on-track) (0.3 pts)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fma_date_logic_048'
SHEET_NAME = 'Tasks'
DATA_ROWS = range(2, 13)   # rows 2 through 12 inclusive (11 rows)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task: cells C2:C12 on sheet 'Tasks' must contain formulas of the form
          =IF(B<row><TODAY(),"Overdue","On Track")
    The initial file has C2:C12 empty — any score above 0 means the agent has acted.
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet 'Tasks' must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"FAIL: Sheet '{SHEET_NAME}' not found. Sheets present: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # -----------------------------------------------------------------------
    # Component 1: At least one cell in C2:C12 contains an IF formula
    #              referencing TODAY() (0.3 points)
    # This FAILS on initial (all C cells are empty) → PASSES on golden
    # -----------------------------------------------------------------------
    cells_with_if_today = []
    for row in DATA_ROWS:
        cell_val = ws.cell(row=row, column=3).value  # column C
        if cell_val and isinstance(cell_val, str):
            val_upper = cell_val.upper().replace(' ', '')
            if 'IF(' in val_upper and 'TODAY()' in val_upper:
                cells_with_if_today.append(f"C{row}")

    try:
        if len(cells_with_if_today) > 0:
            print(f"PASS: Component 1 — {len(cells_with_if_today)} cell(s) in C2:C12 contain IF+TODAY() formula "
                  f"(0.3 pts). First match: {cells_with_if_today[0]}")
            total_score += 0.3
        else:
            print("FAIL: Component 1 — No cells in C2:C12 contain IF formulas with TODAY(). "
                  "Expected =IF(Bx<TODAY(),\"Overdue\",\"On Track\") in each row.")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: ALL 11 cells C2:C12 contain valid IF formulas with TODAY()
    #              (0.4 points)
    # This FAILS on initial → PASSES only if every cell is filled
    # -----------------------------------------------------------------------
    try:
        expected_count = len(list(DATA_ROWS))   # 11
        if len(cells_with_if_today) == expected_count:
            print(f"PASS: Component 2 — All {expected_count} cells C2:C12 contain IF+TODAY() formulas (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 2 — Only {len(cells_with_if_today)}/{expected_count} cells contain "
                  f"IF+TODAY() formulas. Missing rows: "
                  f"{[r for r in DATA_ROWS if f'C{r}' not in cells_with_if_today]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Every formula uses the correct pattern:
    #   =IF(B<row><TODAY(),"Overdue","On Track")
    #   i.e. references the SAME ROW's column B, uses "<" comparison,
    #        and has "Overdue" and "On Track" as the two outcomes.
    #   (0.3 points — only awarded if Component 2 also passed)
    # -----------------------------------------------------------------------
    try:
        correct_formula_count = 0
        formula_issues = []

        # Regex: =IF( B<digits> <TODAY(), "Overdue", "On Track" )
        # Allow case-insensitive matching; spaces allowed inside
        pattern = re.compile(
            r'^\s*=\s*IF\s*\(\s*B(\d+)\s*<\s*TODAY\s*\(\s*\)\s*,'
            r'\s*"Overdue"\s*,\s*"On Track"\s*\)\s*$',
            re.IGNORECASE
        )

        for row in DATA_ROWS:
            cell_val = ws.cell(row=row, column=3).value
            if cell_val and isinstance(cell_val, str):
                m = pattern.match(cell_val)
                if m:
                    ref_row = int(m.group(1))
                    if ref_row == row:
                        correct_formula_count += 1
                    else:
                        formula_issues.append(
                            f"C{row}: formula references B{ref_row} but should reference B{row}"
                        )
                else:
                    formula_issues.append(
                        f"C{row}: formula '{cell_val}' does not match expected pattern "
                        f"=IF(B{row}<TODAY(),\"Overdue\",\"On Track\")"
                    )
            else:
                formula_issues.append(f"C{row}: empty or non-string value: {repr(cell_val)}")

        expected_count = len(list(DATA_ROWS))
        if correct_formula_count == expected_count:
            print(f"PASS: Component 3 — All {expected_count} formulas match correct pattern "
                  f"=IF(Bx<TODAY(),\"Overdue\",\"On Track\") with correct row references (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 3 — Only {correct_formula_count}/{expected_count} formulas "
                  f"have correct structure. Issues:")
            for issue in formula_issues[:5]:  # show first 5 issues
                print(f"  - {issue}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
