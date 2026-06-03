"""
Initial Setup: Task due date status tracker spreadsheet
Task ID: calc_fma_date_logic_048
Domain: libreoffice_calc

Creates a 'Tasks' sheet with:
- Column A (A1:A12): "Task Name" header + 11 realistic task names
- Column B (B1:B12): "Due Date" header + 11 due dates (mix of past and future)
- Column C (C1:C12): "Status" header + C2:C12 EMPTY (awaiting formulas)
"""

import os
from datetime import date, timedelta
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fma_date_logic_048'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Tasks'

    # --- Headers (row 1) ---
    ws['A1'] = 'Task Name'
    ws['B1'] = 'Due Date'
    ws['C1'] = 'Status'

    # --- Task data: mix of past (overdue) and future (on track) dates ---
    # Reference: today is roughly 2026-03-04
    # Past dates (overdue): before today
    # Future dates (on track): after today
    tasks = [
        ('Q1 Financial Report Review',       date(2025, 11, 30)),   # past
        ('Website Redesign Launch',           date(2025, 12, 15)),   # past
        ('Staff Performance Evaluations',     date(2026, 1, 10)),    # past
        ('Vendor Contract Renewal',           date(2026, 2, 28)),    # past
        ('IT Security Audit',                 date(2026, 2, 14)),    # past
        ('Product Roadmap Presentation',      date(2026, 4, 15)),    # future
        ('Client Onboarding Documentation',   date(2026, 5, 1)),     # future
        ('Quarterly Budget Forecast',         date(2026, 3, 20)),    # future
        ('Team Training Workshop',            date(2026, 6, 10)),    # future
        ('Annual Marketing Strategy Review',  date(2026, 7, 31)),    # future
        ('Infrastructure Upgrade Proposal',   date(2026, 3, 12)),    # future (near future)
    ]

    for row_idx, (task_name, due_date) in enumerate(tasks, start=2):
        ws.cell(row=row_idx, column=1, value=task_name)
        ws.cell(row=row_idx, column=2, value=due_date)
        # Column C (Status) is intentionally left empty — awaiting agent formulas

    # Set column widths for readability
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12

    # Apply date number format to column B
    for row_idx in range(2, 13):
        ws.cell(row=row_idx, column=2).number_format = 'yyyy-mm-dd'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

create_initial()
