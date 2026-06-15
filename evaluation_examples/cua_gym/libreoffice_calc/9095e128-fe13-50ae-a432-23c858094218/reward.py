"""
Reward Script: Fill Cost Tier with VLOOKUP and create pivot summary in Sheet2
Task ID: osworld_calc_vlookup_pivot_combined_012
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.5 pts): VLOOKUP formulas in Shipments!C2:C21 (all 20 rows)
  - Component 2 (0.3 pts): Summary sheet has correct total shipping costs by tier
  - Component 3 (0.2 pts): Summary sheet has proper structure (header, tier rows, grand total)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_vlookup_pivot_combined_012'

# Expected pivot values from task requirements
EXPECTED_TIER_TOTALS = {
    'Economy': 94.4,
    'Standard': 245.8,
    'Express': 326.25,
    'Freight': 585.4,
}
EXPECTED_GRAND_TOTAL = 1251.85
TOLERANCE = 0.5  # allow small floating point differences


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: required sheets must exist
    if 'Shipments' not in wb.sheetnames:
        print("CRITICAL: 'Shipments' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    if 'Summary' not in wb.sheetnames:
        print("CRITICAL: 'Summary' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws_shipments = wb['Shipments']
    ws_summary = wb['Summary']

    # -------------------------------------------------------------------------
    # Component 1: VLOOKUP formulas in Shipments!C2:C21 (0.5 points)
    # In the initial file, column C is empty. The task requires filling
    # C2:C21 with VLOOKUP formulas using approximate match against the
    # pricing table in F:G.
    # -------------------------------------------------------------------------
    try:
        vlookup_count = 0
        total_rows = 20  # rows 2 through 21

        for row_idx in range(2, 22):
            cell_val = ws_shipments.cell(row=row_idx, column=3).value
            if cell_val is not None and isinstance(cell_val, str):
                val_upper = cell_val.upper().replace(' ', '')
                # Must be a VLOOKUP formula referencing column B and the pricing table
                if 'VLOOKUP' in val_upper and '$F$' in cell_val and '$G$' in cell_val:
                    vlookup_count += 1

        if vlookup_count == total_rows:
            print(f"PASS: Component 1 — All {total_rows} rows have VLOOKUP formulas in column C (0.5 pts)")
            total_score += 0.5
        elif vlookup_count >= total_rows // 2:
            # Partial credit if at least half the rows are filled
            partial = 0.25
            print(f"PARTIAL: Component 1 — {vlookup_count}/{total_rows} rows have VLOOKUP formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {vlookup_count}/{total_rows} rows have VLOOKUP formulas in column C (expected all 20)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Summary sheet has correct total shipping costs by tier (0.3 points)
    # The task requires Sheet2 (Summary) to show total Shipping Cost grouped by Cost Tier.
    # Expected values: Economy=94.4, Standard=245.8, Express=326.25, Freight=585.4
    # -------------------------------------------------------------------------
    try:
        # Scan all cells in Summary sheet for tier name + value pairs
        found_tiers = {}
        for row in ws_summary.iter_rows(min_row=1, max_row=ws_summary.max_row,
                                         min_col=1, max_col=ws_summary.max_column):
            for i in range(len(row) - 1):
                cell_label = row[i].value
                cell_value = row[i + 1].value
                if cell_label is not None and isinstance(cell_label, str):
                    tier_name = cell_label.strip()
                    if tier_name in EXPECTED_TIER_TOTALS and cell_value is not None:
                        try:
                            found_tiers[tier_name] = float(cell_value)
                        except (ValueError, TypeError):
                            pass

        matched_tiers = 0
        for tier, expected_val in EXPECTED_TIER_TOTALS.items():
            if tier in found_tiers:
                actual_val = found_tiers[tier]
                if abs(actual_val - expected_val) <= TOLERANCE:
                    matched_tiers += 1
                    print(f"  PASS: {tier} total = {actual_val} (expected {expected_val})")
                else:
                    print(f"  FAIL: {tier} total = {actual_val} (expected {expected_val})")
            else:
                print(f"  FAIL: {tier} not found in Summary sheet")

        if matched_tiers == len(EXPECTED_TIER_TOTALS):
            print(f"PASS: Component 2 — All 4 tier totals are correct (0.3 pts)")
            total_score += 0.3
        elif matched_tiers >= 2:
            partial = 0.15
            print(f"PARTIAL: Component 2 — {matched_tiers}/4 tier totals correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {matched_tiers}/4 tier totals found/correct in Summary sheet")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Summary sheet has proper structure (0.2 points)
    # Must have: a header row, at least 4 tier rows, and a Grand Total row.
    # This checks that the summary is not trivially empty or malformed.
    # -------------------------------------------------------------------------
    try:
        header_count = 0   # count of cells containing tier/header keyword
        grand_total_count = 0  # count of cells containing grand total keyword
        tier_rows_found = 0

        for row in ws_summary.iter_rows(min_row=1, max_row=ws_summary.max_row,
                                         min_col=1, max_col=ws_summary.max_column):
            for cell in row:
                if cell.value is not None:
                    val_str = str(cell.value).strip().lower()
                    if 'cost tier' in val_str or ('tier' in val_str and 'total' not in val_str):
                        header_count += 1
                    if 'grand total' in val_str:
                        grand_total_count += 1
            # Count rows that have a known tier name
            for cell in row:
                if cell.value is not None and str(cell.value).strip() in EXPECTED_TIER_TOTALS:
                    tier_rows_found += 1

        if header_count >= 1 and grand_total_count >= 1 and tier_rows_found >= 4:
            print(f"PASS: Component 3 — Summary sheet has header, {tier_rows_found} tier rows, and grand total (0.2 pts)")
            total_score += 0.2
        elif tier_rows_found >= 2:
            partial = 0.1
            print(f"PARTIAL: Component 3 — Summary sheet has {tier_rows_found} tier rows but missing header or grand total ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Summary sheet structure invalid: header_count={header_count}, grand_total_count={grand_total_count}, tier_rows={tier_rows_found}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
