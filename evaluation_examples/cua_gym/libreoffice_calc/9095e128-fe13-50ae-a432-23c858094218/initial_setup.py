"""
Initial Setup: Fill shipping cost tier via VLOOKUP and create pivot in Sheet2
Task ID: osworld_calc_vlookup_pivot_combined_012
Domain: libreoffice_calc

Creates Sheet1 with shipping records (Shipment ID, Weight (kg), Cost Tier [empty], Shipping Cost)
plus a weight-based pricing table in columns F-G.
Sheet2 is created empty, ready for the agent to build a pivot table.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_vlookup_pivot_combined_012'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: Shipments ---
    ws1 = wb.active
    ws1.title = 'Shipments'

    # Header row styling
    header_font = Font(bold=True, name='Calibri', size=11)
    header_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')

    # Main table headers (A-D)
    main_headers = ['Shipment ID', 'Weight (kg)', 'Cost Tier', 'Shipping Cost']
    for col, h in enumerate(main_headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Pricing table headers (F-G)
    pricing_headers = ['Min Weight (kg)', 'Tier']
    for col, h in enumerate(pricing_headers, 6):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color='FFFFF2CC', end_color='FFFFF2CC', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Weight-based pricing table in F-G (approximate match VLOOKUP requires sorted ascending)
    # Tier thresholds: 0-4.99 = Economy, 5-14.99 = Standard, 15-29.99 = Express, 30+ = Freight
    pricing_table = [
        (0,   'Economy'),
        (5,   'Standard'),
        (15,  'Express'),
        (30,  'Freight'),
    ]
    for r, (min_wt, tier) in enumerate(pricing_table, 2):
        ws1.cell(row=r, column=6, value=min_wt)
        ws1.cell(row=r, column=7, value=tier)

    # Shipment data rows (realistic content, Cost Tier column C left EMPTY)
    shipment_data = [
        ('SHP-1001', 2.3,  None, 18.50),
        ('SHP-1002', 12.7, None, 45.00),
        ('SHP-1003', 0.8,  None, 9.75),
        ('SHP-1004', 22.5, None, 78.20),
        ('SHP-1005', 6.1,  None, 32.40),
        ('SHP-1006', 45.0, None, 145.00),
        ('SHP-1007', 3.4,  None, 21.60),
        ('SHP-1008', 18.9, None, 65.30),
        ('SHP-1009', 9.2,  None, 38.90),
        ('SHP-1010', 31.5, None, 112.00),
        ('SHP-1011', 1.5,  None, 12.25),
        ('SHP-1012', 28.4, None, 95.75),
        ('SHP-1013', 7.8,  None, 35.60),
        ('SHP-1014', 52.3, None, 198.40),
        ('SHP-1015', 4.1,  None, 24.80),
        ('SHP-1016', 14.2, None, 51.15),
        ('SHP-1017', 0.4,  None, 7.50),
        ('SHP-1018', 38.7, None, 130.00),
        ('SHP-1019', 11.6, None, 42.75),
        ('SHP-1020', 25.0, None, 87.00),
    ]

    for r, (ship_id, weight, tier, cost) in enumerate(shipment_data, 2):
        ws1.cell(row=r, column=1, value=ship_id)
        ws1.cell(row=r, column=2, value=weight)
        # Column C (Cost Tier) intentionally left empty
        ws1.cell(row=r, column=4, value=cost)
        ws1.cell(row=r, column=4).number_format = '#,##0.00'

    # Column widths
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 14
    ws1.column_dimensions['C'].width = 14
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 4   # spacer
    ws1.column_dimensions['F'].width = 18
    ws1.column_dimensions['G'].width = 12

    # --- Sheet2: Summary (empty, agent will create pivot here) ---
    ws2 = wb.create_sheet('Summary')
    ws2.cell(row=1, column=1, value='')  # placeholder

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
