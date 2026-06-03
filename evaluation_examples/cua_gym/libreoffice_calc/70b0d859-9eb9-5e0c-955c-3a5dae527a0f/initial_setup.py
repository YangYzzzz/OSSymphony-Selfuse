"""
Initial Setup: Loan application data with credit score reference table for VLOOKUP + pivot table task
Task ID: osworld_calc_vlookup_pivot_combined_015
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_vlookup_pivot_combined_015'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: LoanApplications ---
    ws1 = wb.active
    ws1.title = 'LoanApplications'

    # Header style
    header_font = Font(bold=True, name='Calibri', size=11)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    white_font = Font(bold=True, name='Calibri', size=11, color='FFFFFFFF')

    # Main table headers: A1:D1
    main_headers = ['Application ID', 'Credit Score', 'Risk Rating', 'Loan Amount']
    for col, h in enumerate(main_headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Loan data (Risk Rating column intentionally EMPTY — task must fill it via VLOOKUP)
    loan_data = [
        ('APP-001', 720, None, 45000),
        ('APP-002', 580, None, 12000),
        ('APP-003', 640, None, 25000),
        ('APP-004', 790, None, 60000),
        ('APP-005', 520, None, 8000),
        ('APP-006', 680, None, 30000),
        ('APP-007', 750, None, 55000),
        ('APP-008', 350, None, 5000),
        ('APP-009', 710, None, 42000),
        ('APP-010', 620, None, 18000),
        ('APP-011', 490, None, 9500),
        ('APP-012', 760, None, 70000),
        ('APP-013', 590, None, 15000),
        ('APP-014', 810, None, 80000),
        ('APP-015', 650, None, 22000),
    ]

    for r, (app_id, score, rating, amount) in enumerate(loan_data, 2):
        ws1.cell(row=r, column=1, value=app_id)
        ws1.cell(row=r, column=2, value=score)
        # Column C (Risk Rating) intentionally left blank — agent must fill with VLOOKUP
        ws1.cell(row=r, column=4, value=amount)

    # Credit Score Reference Table in columns F-G
    # This is the lookup table for VLOOKUP (sorted ascending for approximate match)
    ref_header_fill = PatternFill(start_color='FF70AD47', end_color='FF70AD47', fill_type='solid')

    ws1.cell(row=1, column=6, value='Min Score').font = Font(bold=True, name='Calibri', size=11, color='FFFFFFFF')
    ws1.cell(row=1, column=6).fill = ref_header_fill
    ws1.cell(row=1, column=6).alignment = Alignment(horizontal='center')

    ws1.cell(row=1, column=7, value='Risk Rating').font = Font(bold=True, name='Calibri', size=11, color='FFFFFFFF')
    ws1.cell(row=1, column=7).fill = ref_header_fill
    ws1.cell(row=1, column=7).alignment = Alignment(horizontal='center')

    # Reference data: sorted ascending by score for approximate match VLOOKUP
    ref_data = [
        (300, 'High'),    # Score 300-599 → High
        (600, 'Medium'),  # Score 600-699 → Medium
        (700, 'Low'),     # Score 700+    → Low
    ]
    for r, (score_threshold, risk) in enumerate(ref_data, 2):
        ws1.cell(row=r, column=6, value=score_threshold)
        ws1.cell(row=r, column=7, value=risk)

    # Column widths
    ws1.column_dimensions['A'].width = 16
    ws1.column_dimensions['B'].width = 14
    ws1.column_dimensions['C'].width = 14
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 4
    ws1.column_dimensions['F'].width = 12
    ws1.column_dimensions['G'].width = 12

    # Freeze header row
    ws1.freeze_panes = 'A2'

    # --- Sheet 2: PivotSummary (empty — agent must create pivot table here) ---
    ws2 = wb.create_sheet('PivotSummary')
    # Leave empty so the agent can create the pivot table

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
