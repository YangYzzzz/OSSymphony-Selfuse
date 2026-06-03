"""
Reward Script: Fill Risk Rating using VLOOKUP and create pivot table in Sheet2
Task ID: osworld_calc_vlookup_pivot_combined_015
Domain: libreoffice_calc
Scoring:
  Component 1 (0.50): VLOOKUP formulas present in Risk Rating column (C2:C16),
                       referencing credit score table with approximate match
  Component 2 (0.30): PivotSummary sheet has at least 3 risk category data rows
  Component 3 (0.20): PivotSummary totals are numerically correct
                       (High=49500, Low=352000, Medium=95000)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_vlookup_pivot_combined_015'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook to read formula strings
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: LoanApplications sheet must exist
    if 'LoanApplications' not in wb.sheetnames:
        print("FAIL: 'LoanApplications' sheet not found in workbook")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws_loans = wb['LoanApplications']

    # -------------------------------------------------------------------
    # Component 1: VLOOKUP formulas in Risk Rating column C (rows 2-16)
    #   (0.5 points)
    #
    #   The task asks the agent to fill in Risk Rating (column C) using
    #   VLOOKUP with approximate match (last arg = 1 or TRUE) referencing
    #   the credit score reference table in columns F-G.
    #
    #   Scoring breakdown:
    #   - Full 0.5: >= 14/15 cells have VLOOKUP with approx match + F/G reference
    #   - Partial 0.3: VLOOKUP + approx match but wrong table reference
    #   - Partial 0.2: VLOOKUP present but exact match (last arg=0/FALSE)
    # -------------------------------------------------------------------
    try:
        vlookup_count = 0
        approx_match_count = 0
        ref_table_count = 0

        for row_idx in range(2, 17):  # rows 2 to 16 (15 data rows)
            cell = ws_loans.cell(row=row_idx, column=3)
            cell_val = cell.value
            if isinstance(cell_val, str):
                # Check it is a VLOOKUP formula
                if cell_val.upper().replace(' ', '').startswith('=VLOOKUP('):
                    vlookup_count += 1
                    # Check approximate match: last arg is 1 or TRUE
                    if re.search(r',\s*1\s*\)', cell_val) or re.search(r',\s*TRUE\s*\)', cell_val, re.IGNORECASE):
                        approx_match_count += 1
                    # Check it references the F/G credit score reference table
                    if '$F$' in cell_val or '$F' in cell_val or 'F2:G' in cell_val.upper():
                        ref_table_count += 1

        # Award points based on quality of VLOOKUP implementation
        if vlookup_count >= 14 and approx_match_count >= 14 and ref_table_count >= 14:
            print(f"PASS: Component 1 — {vlookup_count}/15 VLOOKUP formulas, "
                  f"{approx_match_count} approx match, {ref_table_count} ref F/G table")
            total_score += 0.5
        elif vlookup_count >= 14 and approx_match_count >= 14:
            print(f"PARTIAL: Component 1 — VLOOKUP + approx match but only {ref_table_count}/15 reference F/G table (0.3 pts)")
            total_score += 0.3
        elif vlookup_count >= 14:
            print(f"PARTIAL: Component 1 — {vlookup_count}/15 VLOOKUP formulas but only {approx_match_count} use approx match (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 1 — only {vlookup_count}/15 Risk Rating cells have VLOOKUP formulas (need >= 14)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------
    # Component 2: PivotSummary sheet has summary data rows
    #   (0.3 points)
    #
    #   The task asks for a pivot table in Sheet2 (named PivotSummary)
    #   summarizing total Loan Amount by Risk Rating.
    #   We verify:
    #   (a) 'PivotSummary' sheet exists (or Sheet2/Summary as fallback)
    #   (b) It has at least 3 data rows (Low, Medium, High categories)
    # -------------------------------------------------------------------
    try:
        # Find the pivot summary sheet — try PivotSummary first, then fallbacks
        ws_pivot = None
        if 'PivotSummary' in wb.sheetnames:
            ws_pivot = wb['PivotSummary']
        else:
            for sname in wb.sheetnames:
                if sname.lower() in ('sheet2', 'pivot', 'summary'):
                    ws_pivot = wb[sname]
                    break

        if ws_pivot is None:
            print("FAIL: Component 2 — No PivotSummary sheet found in workbook")
        else:
            data_rows = 0
            risk_ratings_found = set()
            for row in ws_pivot.iter_rows(min_row=2, max_row=ws_pivot.max_row, values_only=True):
                if row[0] is not None and str(row[0]).strip() not in ('', 'Grand Total', 'Total'):
                    data_rows += 1
                    risk_ratings_found.add(str(row[0]).strip())

            if data_rows >= 3:
                print(f"PASS: Component 2 — {ws_pivot.title} has {data_rows} risk category rows: {sorted(risk_ratings_found)}")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — {ws_pivot.title} has only {data_rows} category rows (need >= 3, found: {risk_ratings_found})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------
    # Component 3: PivotSummary total amounts are numerically correct
    #   (0.2 points)
    #
    #   Expected totals (based on VLOOKUP mapping: 300+=High, 600+=Medium, 700+=Low):
    #   - High   (score 300-599): 49500
    #   - Medium (score 600-699): 95000
    #   - Low    (score 700+):   352000
    #   - Grand Total:           496500
    #
    #   Full 0.2 if all 3 categories are correct; partial 0.1 if >= 2 correct
    # -------------------------------------------------------------------
    try:
        # Find pivot sheet
        ws_pivot2 = None
        if 'PivotSummary' in wb.sheetnames:
            ws_pivot2 = wb['PivotSummary']
        else:
            for sname in wb.sheetnames:
                if sname.lower() in ('sheet2', 'pivot', 'summary'):
                    ws_pivot2 = wb[sname]
                    break

        if ws_pivot2 is None:
            print("FAIL: Component 3 — No pivot summary sheet found")
        else:
            # Extract (label -> amount) mapping from the pivot sheet
            label_to_amount = {}
            for row in ws_pivot2.iter_rows(min_row=2, max_row=ws_pivot2.max_row, values_only=True):
                if row[0] is not None and row[1] is not None:
                    label = str(row[0]).strip()
                    try:
                        amount = float(row[1])
                        label_to_amount[label] = amount
                    except (ValueError, TypeError):
                        pass

            # Expected values derived from task data
            EXPECTED = {
                'High': 49500.0,
                'Low': 352000.0,
                'Medium': 95000.0,
            }
            TOLERANCE = 1.0  # allow small rounding differences

            correct_count = 0
            for risk, expected_amount in EXPECTED.items():
                if risk in label_to_amount:
                    actual = label_to_amount[risk]
                    if abs(actual - expected_amount) <= TOLERANCE:
                        correct_count += 1
                        print(f"PASS: Component 3 — {risk} total = {actual} (expected {expected_amount})")
                    else:
                        print(f"FAIL: Component 3 — {risk} total = {actual} (expected {expected_amount})")
                else:
                    print(f"FAIL: Component 3 — Risk category '{risk}' not found in pivot summary")

            if correct_count == 3:
                print("PASS: Component 3 — All 3 risk category totals are correct (0.2 pts)")
                total_score += 0.2
            elif correct_count >= 2:
                print(f"PARTIAL: Component 3 — {correct_count}/3 category totals correct (partial credit 0.1 pts)")
                total_score += 0.1
            else:
                print(f"FAIL: Component 3 — Only {correct_count}/3 category totals are correct")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
