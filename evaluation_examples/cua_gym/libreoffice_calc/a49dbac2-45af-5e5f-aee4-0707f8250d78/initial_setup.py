"""
Initial Setup: Create a loan calculation spreadsheet with interest rate in B1.
Task ID: calc_nrv_009
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_009'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Loan Parameters ---
    ws1 = wb.active
    ws1.title = 'Loan Parameters'

    # Header styling
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    # Row 1: Interest Rate parameter
    ws1['A1'] = 'Annual Interest Rate'
    ws1['A1'].font = Font(name='Arial', size=11, bold=True)
    ws1['B1'] = 0.045
    ws1['B1'].number_format = '0.00%'
    ws1['B1'].font = Font(name='Arial', size=11)

    # Row 2: Blank separator
    # Row 3: Loan table headers
    headers = ['Borrower', 'Loan Amount', 'Term (Years)', 'Monthly Payment', 'Total Interest', 'Total Paid']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Loan data - realistic borrowers
    loans = [
        ['Sarah Chen', 250000, 30],
        ['Marcus Johnson', 185000, 15],
        ['Elena Rodriguez', 320000, 30],
        ['David Kim', 150000, 20],
        ['Rachel Thompson', 275000, 25],
        ['James O\'Brien', 410000, 30],
        ['Aisha Patel', 198000, 15],
        ['Michael Torres', 225000, 20],
        ['Lisa Wang', 340000, 30],
        ['Robert Fischer', 175000, 25],
        ['Amanda Nguyen', 290000, 15],
        ['Christopher Lee', 215000, 20],
    ]

    data_font = Font(name='Arial', size=11)
    currency_fmt = '$#,##0.00'
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    for r, (name, amount, term) in enumerate(loans, 4):
        ws1.cell(row=r, column=1, value=name).font = data_font
        ws1.cell(row=r, column=2, value=amount).font = data_font
        ws1.cell(row=r, column=2).number_format = currency_fmt
        ws1.cell(row=r, column=3, value=term).font = data_font

        # Monthly Payment formula: =PMT(B1/12, C{r}*12, -B{r})
        # References B1 directly (no named range yet)
        ws1.cell(row=r, column=4, value=f'=PMT(B$1/12,C{r}*12,-B{r})').font = data_font
        ws1.cell(row=r, column=4).number_format = currency_fmt

        # Total Interest: =D{r}*C{r}*12 - B{r}
        ws1.cell(row=r, column=5, value=f'=D{r}*C{r}*12-B{r}').font = data_font
        ws1.cell(row=r, column=5).number_format = currency_fmt

        # Total Paid: =D{r}*C{r}*12
        ws1.cell(row=r, column=6, value=f'=D{r}*C{r}*12').font = data_font
        ws1.cell(row=r, column=6).number_format = currency_fmt

        # Borders for all cells in row
        for c in range(1, 7):
            ws1.cell(row=r, column=c).border = thin_border

    # Column widths
    ws1.column_dimensions['A'].width = 22
    ws1.column_dimensions['B'].width = 16
    ws1.column_dimensions['C'].width = 14
    ws1.column_dimensions['D'].width = 18
    ws1.column_dimensions['E'].width = 18
    ws1.column_dimensions['F'].width = 16

    # --- Sheet 2: Amortization Schedule (for first loan) ---
    ws2 = wb.create_sheet('Amortization')

    ws2['A1'] = 'Amortization Schedule - Sarah Chen'
    ws2['A1'].font = Font(name='Arial', size=13, bold=True)

    am_headers = ['Month', 'Beginning Balance', 'Payment', 'Principal', 'Interest', 'Ending Balance']
    for col, h in enumerate(am_headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # First 12 months of amortization with formulas referencing B1 on Loan Parameters
    for month in range(1, 13):
        r = month + 3
        ws2.cell(row=r, column=1, value=month).font = data_font

        if month == 1:
            # Beginning balance = loan amount from Loan Parameters
            ws2.cell(row=r, column=2, value=f"='Loan Parameters'!B4").font = data_font
        else:
            # Beginning balance = previous ending balance
            ws2.cell(row=r, column=2, value=f'=F{r-1}').font = data_font

        # Payment (constant monthly)
        ws2.cell(row=r, column=3, value=f"=PMT('Loan Parameters'!B$1/12,'Loan Parameters'!C4*12,-'Loan Parameters'!B4)").font = data_font

        # Interest portion = beginning balance * monthly rate (references B1 on Loan Parameters)
        ws2.cell(row=r, column=5, value=f"=B{r}*'Loan Parameters'!B$1/12").font = data_font

        # Principal = payment - interest
        ws2.cell(row=r, column=4, value=f'=C{r}-E{r}').font = data_font

        # Ending balance = beginning - principal
        ws2.cell(row=r, column=6, value=f'=B{r}-D{r}').font = data_font

        for c in range(1, 7):
            ws2.cell(row=r, column=c).number_format = currency_fmt if c > 1 else '0'
            ws2.cell(row=r, column=c).border = thin_border

    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 16
    ws2.column_dimensions['D'].width = 16
    ws2.column_dimensions['E'].width = 16
    ws2.column_dimensions['F'].width = 20

    # --- Sheet 3: Summary ---
    ws3 = wb.create_sheet('Rate Comparison')

    ws3['A1'] = 'Interest Rate Sensitivity Analysis'
    ws3['A1'].font = Font(name='Arial', size=13, bold=True)

    ws3['A3'] = 'Current Rate:'
    ws3['A3'].font = Font(name='Arial', size=11, bold=True)
    ws3['B3'] = "='Loan Parameters'!B1"
    ws3['B3'].number_format = '0.00%'

    comp_headers = ['Rate Scenario', 'Rate', 'Monthly Payment (250K/30yr)', 'Difference from Current']
    for col, h in enumerate(comp_headers, 1):
        cell = ws3.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    scenarios = [
        ['Current Rate', "='Loan Parameters'!B1"],
        ['Down 0.5%', "='Loan Parameters'!B1-0.005"],
        ['Down 1.0%', "='Loan Parameters'!B1-0.01"],
        ['Up 0.5%', "='Loan Parameters'!B1+0.005"],
        ['Up 1.0%', "='Loan Parameters'!B1+0.01"],
    ]

    for r, (label, rate_formula) in enumerate(scenarios, 6):
        ws3.cell(row=r, column=1, value=label).font = data_font
        ws3.cell(row=r, column=2, value=rate_formula).font = data_font
        ws3.cell(row=r, column=2).number_format = '0.00%'
        # Monthly payment for $250K, 30 years at this rate
        ws3.cell(row=r, column=3, value=f'=PMT(B{r}/12,360,-250000)').font = data_font
        ws3.cell(row=r, column=3).number_format = currency_fmt
        # Difference from current
        ws3.cell(row=r, column=4, value=f'=C{r}-C$6').font = data_font
        ws3.cell(row=r, column=4).number_format = currency_fmt

    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 12
    ws3.column_dimensions['C'].width = 30
    ws3.column_dimensions['D'].width = 26

    # NO named ranges - that's what the task is about
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
