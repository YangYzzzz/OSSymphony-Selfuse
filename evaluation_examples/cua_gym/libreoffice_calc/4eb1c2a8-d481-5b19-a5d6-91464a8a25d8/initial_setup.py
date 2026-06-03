"""
Initial Setup: Create spreadsheet with Record Number, Name, and Code columns.
Task ID: calc_gcv_061
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_061'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Unique_Codes"

    # --- Headers ---
    headers = ['Record Number', 'Name', 'Code']
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    # --- Data rows (19 rows: rows 2-20) ---
    names = [
        'Sarah Chen', 'Marcus Johnson', 'Priya Patel', 'James O\'Brien',
        'Yuki Tanaka', 'Elena Rodriguez', 'David Kim', 'Fatima Al-Hassan',
        'Lucas Bergmann', 'Aisha Williams', 'Roberto Silva', 'Mei Lin Zhang',
        'Thomas Wright', 'Ingrid Svensson', 'Omar Farouk', 'Sophie Dubois',
        'Kenji Watanabe', 'Amara Okafor', 'Henrik Larsen'
    ]

    # Some codes already entered (12 out of 19), some blank
    codes = [
        'UC-1001', 'UC-1002', 'UC-1003', None,
        'UC-1005', None, 'UC-1007', 'UC-1008',
        None, 'UC-1010', 'UC-1011', None,
        'UC-1013', 'UC-1014', None, 'UC-1016',
        None, 'UC-1018', 'UC-1019'
    ]

    for i in range(19):
        row = i + 2
        ws.cell(row=row, column=1, value=i + 1)  # Record Number
        ws.cell(row=row, column=2, value=names[i])  # Name
        if codes[i] is not None:
            ws.cell(row=row, column=3, value=codes[i])  # Code

    # --- Column widths ---
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 16

    # No data validation on initial file

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
