"""
Reward Script: Track production line changeover times — calculate duration, over-target, lost
production units, and populate summary AVERAGEIFS/COUNTIFS formulas plus Pareto bar chart.
Task ID: calc_ops_production_changeover_tracking_068
Domain: libreoffice_calc
Scoring:
  - Component 1: ChangeoverLog G column formulas (Duration Minutes) — 0.30 pts
  - Component 2: ChangeoverLog I column formulas (Over Target)       — 0.25 pts
  - Component 3: ChangeoverLog J column formulas (Lost Production)   — 0.20 pts
  - Component 4: ChangeoverSummary AVERAGEIFS/COUNTIFS formulas      — 0.15 pts
  - Component 5: Pareto bar chart in ChangeoverSummary               — 0.10 pts
Total: 1.0
"""

import os
import openpyxl
from openpyxl.utils import get_column_letter

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_production_changeover_tracking_068'


def normalize_formula(f):
    """Normalize a formula string for comparison: uppercase, no spaces."""
    if f is None:
        return ''
    return str(f).upper().replace(' ', '')


def check_g_formula(ws, row):
    """Check if G<row> contains formula =(F<row>-E<row>)*1440."""
    val = ws.cell(row=row, column=7).value  # Column G: Duration Minutes
    if val is None:
        return False
    norm = normalize_formula(val)
    expected = '=(F{r}-E{r})*1440'.format(r=row).upper()
    return norm == expected


def check_i_formula(ws, row):
    """Check if I<row> contains formula =MAX(0,G<row>-H<row>)."""
    val = ws.cell(row=row, column=9).value  # Column I: Over Target
    if val is None:
        return False
    norm = normalize_formula(val)
    expected = '=MAX(0,G{r}-H{r})'.format(r=row).upper()
    return norm == expected


def check_j_formula(ws, row):
    """Check if J<row> contains formula =I<row>/60*K<row>."""
    val = ws.cell(row=row, column=10).value  # Column J: Lost Production Units
    if val is None:
        return False
    norm = normalize_formula(val)
    expected = '=I{r}/60*K{r}'.format(r=row).upper()
    return norm == expected


def check_summary_formula(ws, row, col):
    """
    Check ChangeoverSummary formula.
    col=2 (B): expects COUNTIFS referencing ChangeoverLog.
    col=3,4,5 (C,D,E): expects AVERAGEIFS referencing ChangeoverLog.
    """
    val = ws.cell(row=row, column=col).value
    if val is None:
        return False
    norm = normalize_formula(val)
    if col == 2:
        return 'COUNTIFS' in norm and 'CHANGEOVERLOG' in norm
    return 'AVERAGEIFS' in norm and 'CHANGEOVERLOG' in norm


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print('CRITICAL: Cannot load file ' + str(file_path) + ': ' + str(e))
        print('REWARD: 0.0')
        return 0.0

    # Precondition gate — required sheets must exist (not scored)
    if 'ChangeoverLog' not in wb.sheetnames or 'ChangeoverSummary' not in wb.sheetnames:
        print('CRITICAL: Required sheets missing. Found: ' + str(wb.sheetnames))
        print('REWARD: 0.0')
        return 0.0

    ws_log = wb['ChangeoverLog']
    ws_sum = wb['ChangeoverSummary']

    # Component 1: ChangeoverLog G column (Duration Minutes = (End-Start)*1440) — 0.30 pts
    # G2:G41 must all have formulas of the form =(Fx-Ex)*1440
    try:
        g_pass = sum(1 for r in range(2, 42) if check_g_formula(ws_log, r))
        g_fail_rows = [r for r in range(2, 42) if not check_g_formula(ws_log, r)]

        if g_pass == 40:
            total_score += 0.30
            print('PASS: Component 1 — All 40 G-col Duration formulas present (=(Fx-Ex)*1440) (0.30 pts)')
        elif g_pass >= 30:
            partial = round(0.30 * g_pass / 40, 4)
            total_score += partial
            print('PARTIAL: Component 1 — ' + str(g_pass) + '/40 G-col Duration formulas present (' + str(partial) + ' pts)')
            print('  Failed rows (sample): ' + str(g_fail_rows[:5]))
        else:
            print('FAIL: Component 1 — Only ' + str(g_pass) + '/40 G-col Duration formulas present')
            print('  Failed rows (sample): ' + str(g_fail_rows[:5]))
            print('  G2 value: ' + repr(ws_log.cell(row=2, column=7).value))
    except Exception as e:
        print('ERROR: Component 1 — ' + str(e))

    # Component 2: ChangeoverLog I column (Over Target = MAX(0, G-H)) — 0.25 pts
    # I2:I41 must all have formulas of the form =MAX(0,Gx-Hx)
    try:
        i_pass = sum(1 for r in range(2, 42) if check_i_formula(ws_log, r))
        i_fail_rows = [r for r in range(2, 42) if not check_i_formula(ws_log, r)]

        if i_pass == 40:
            total_score += 0.25
            print('PASS: Component 2 — All 40 I-col Over Target formulas present (=MAX(0,Gx-Hx)) (0.25 pts)')
        elif i_pass >= 30:
            partial = round(0.25 * i_pass / 40, 4)
            total_score += partial
            print('PARTIAL: Component 2 — ' + str(i_pass) + '/40 I-col Over Target formulas present (' + str(partial) + ' pts)')
            print('  Failed rows (sample): ' + str(i_fail_rows[:5]))
        else:
            print('FAIL: Component 2 — Only ' + str(i_pass) + '/40 I-col Over Target formulas present')
            print('  Failed rows (sample): ' + str(i_fail_rows[:5]))
            print('  I2 value: ' + repr(ws_log.cell(row=2, column=9).value))
    except Exception as e:
        print('ERROR: Component 2 — ' + str(e))

    # Component 3: ChangeoverLog J column (Lost Production Units = I/60*K) — 0.20 pts
    # J2:J41 must all have formulas of the form =Ix/60*Kx
    try:
        j_pass = sum(1 for r in range(2, 42) if check_j_formula(ws_log, r))
        j_fail_rows = [r for r in range(2, 42) if not check_j_formula(ws_log, r)]

        if j_pass == 40:
            total_score += 0.20
            print('PASS: Component 3 — All 40 J-col Lost Production formulas present (=Ix/60*Kx) (0.20 pts)')
        elif j_pass >= 30:
            partial = round(0.20 * j_pass / 40, 4)
            total_score += partial
            print('PARTIAL: Component 3 — ' + str(j_pass) + '/40 J-col Lost Production formulas present (' + str(partial) + ' pts)')
            print('  Failed rows (sample): ' + str(j_fail_rows[:5]))
        else:
            print('FAIL: Component 3 — Only ' + str(j_pass) + '/40 J-col Lost Production formulas present')
            print('  Failed rows (sample): ' + str(j_fail_rows[:5]))
            print('  J2 value: ' + repr(ws_log.cell(row=2, column=10).value))
    except Exception as e:
        print('ERROR: Component 3 — ' + str(e))

    # Component 4: ChangeoverSummary AVERAGEIFS/COUNTIFS formulas — 0.15 pts
    # Rows 2-11 (10 transitions): col B=COUNTIFS, cols C/D/E=AVERAGEIFS referencing ChangeoverLog
    try:
        summary_pass = 0
        summary_total = 0
        summary_fail_cells = []

        for row in range(2, 12):      # rows 2-11 (10 transitions)
            for col in range(2, 6):   # columns B=2, C=3, D=4, E=5
                summary_total += 1
                if check_summary_formula(ws_sum, row, col):
                    summary_pass += 1
                else:
                    summary_fail_cells.append(get_column_letter(col) + str(row))

        if summary_pass == summary_total:
            total_score += 0.15
            print('PASS: Component 4 — All ' + str(summary_total) + ' ChangeoverSummary formulas present (COUNTIFS/AVERAGEIFS) (0.15 pts)')
        elif summary_pass >= int(summary_total * 0.7):
            partial = round(0.15 * summary_pass / summary_total, 4)
            total_score += partial
            print('PARTIAL: Component 4 — ' + str(summary_pass) + '/' + str(summary_total) + ' ChangeoverSummary formulas (' + str(partial) + ' pts)')
            print('  Failed cells (sample): ' + str(summary_fail_cells[:5]))
        else:
            print('FAIL: Component 4 — Only ' + str(summary_pass) + '/' + str(summary_total) + ' ChangeoverSummary formulas present')
            print('  Failed cells (sample): ' + str(summary_fail_cells[:5]))
            print('  B2 value: ' + repr(ws_sum.cell(row=2, column=2).value))
    except Exception as e:
        print('ERROR: Component 4 — ' + str(e))

    # Component 5: Pareto bar chart in ChangeoverSummary — 0.10 pts
    # Task requires a bar chart (sorted by Avg Duration descending) for Pareto analysis
    try:
        charts_sum = ws_sum._charts
        charts_log = ws_log._charts

        # Accept chart in ChangeoverSummary (preferred) or ChangeoverLog
        if len(charts_sum) > 0:
            print('PASS: Component 5 — ' + str(len(charts_sum)) + ' chart(s) found in ChangeoverSummary (Pareto) (0.10 pts)')
            total_score += 0.10
        elif len(charts_log) > 0:
            print('PASS: Component 5 — ' + str(len(charts_log)) + ' chart(s) found in ChangeoverLog (Pareto) (0.10 pts)')
            total_score += 0.10
        else:
            print('FAIL: Component 5 — No bar/Pareto chart found in either sheet')
            print('  ChangeoverSummary charts: ' + str(len(charts_sum)))
            print('  ChangeoverLog charts: ' + str(len(charts_log)))
    except Exception as e:
        print('ERROR: Component 5 — ' + str(e))

    final_score = round(min(total_score, 1.0), 4)
    print('')
    print('Score: ' + str(round(total_score, 4)) + '/1.0')
    print('REWARD: ' + str(final_score))
    return final_score


# Default: test against golden file (path on VM)
file_path = WORKDIR + '/' + TASK_ID + '_initial.xlsx'
if not os.path.exists(file_path):
    print('File not found: ' + file_path)
    print('REWARD: 0.0')
else:
    verify_task(file_path)
