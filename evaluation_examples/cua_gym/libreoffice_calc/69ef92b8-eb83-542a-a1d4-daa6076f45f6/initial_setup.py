"""
Initial Setup: Production Changeover Tracking
Task ID: calc_ops_production_changeover_tracking_068
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, date, timedelta, time
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_production_changeover_tracking_068'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ #
    # Sheet 1: ChangeoverLog
    # ------------------------------------------------------------------ #
    ws1 = wb.active
    ws1.title = 'ChangeoverLog'

    headers = [
        'Date', 'Line', 'From Product', 'To Product',
        'Changeover Start', 'Changeover End',
        'Duration Minutes',   # G — empty
        'Target Minutes',
        'Over Target',        # I — empty
        'Lost Production Units',  # J — empty
        'Units per Hour'
    ]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws1.row_dimensions[1].height = 30
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 10
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 18
    ws1.column_dimensions['E'].width = 18
    ws1.column_dimensions['F'].width = 18
    ws1.column_dimensions['G'].width = 16
    ws1.column_dimensions['H'].width = 14
    ws1.column_dimensions['I'].width = 12
    ws1.column_dimensions['J'].width = 20
    ws1.column_dimensions['K'].width = 14

    # Production lines
    lines = ['Line A', 'Line B', 'Line C', 'Line D']

    # Products
    products = [
        'SKU-1001 Granola Bar 40g',
        'SKU-1002 Protein Bar 55g',
        'SKU-1003 Cereal Box 500g',
        'SKU-1004 Oat Crackers 200g',
        'SKU-1005 Trail Mix 300g',
        'SKU-1006 Rice Cakes 150g',
        'SKU-1007 Muesli 750g',
        'SKU-1008 Nut Butter 250g',
    ]

    # Target changeover minutes by product pair complexity
    # Quick changeovers: same product family ~15-25 min
    # Slow changeovers: different families ~40-60 min
    targets = {
        ('SKU-1001 Granola Bar 40g', 'SKU-1002 Protein Bar 55g'): 20,
        ('SKU-1002 Protein Bar 55g', 'SKU-1001 Granola Bar 40g'): 20,
        ('SKU-1003 Cereal Box 500g', 'SKU-1007 Muesli 750g'): 25,
        ('SKU-1007 Muesli 750g', 'SKU-1003 Cereal Box 500g'): 25,
        ('SKU-1004 Oat Crackers 200g', 'SKU-1006 Rice Cakes 150g'): 30,
        ('SKU-1006 Rice Cakes 150g', 'SKU-1004 Oat Crackers 200g'): 30,
        ('SKU-1005 Trail Mix 300g', 'SKU-1008 Nut Butter 250g'): 35,
        ('SKU-1008 Nut Butter 250g', 'SKU-1005 Trail Mix 300g'): 35,
    }
    default_target = 45

    # Units per hour for each product
    uph = {
        'SKU-1001 Granola Bar 40g': 1800,
        'SKU-1002 Protein Bar 55g': 1500,
        'SKU-1003 Cereal Box 500g': 600,
        'SKU-1004 Oat Crackers 200g': 900,
        'SKU-1005 Trail Mix 300g': 750,
        'SKU-1006 Rice Cakes 150g': 1200,
        'SKU-1007 Muesli 750g': 480,
        'SKU-1008 Nut Butter 250g': 560,
    }

    # Generate 40 changeover records over the past week (Mon-Fri)
    # Week of 2026-02-23 to 2026-02-27
    week_dates = [
        date(2026, 2, 23),
        date(2026, 2, 24),
        date(2026, 2, 25),
        date(2026, 2, 26),
        date(2026, 2, 27),
    ]

    # Predefined records for determinism and realistic variety
    records = [
        # (date_idx, line, from_prod, to_prod, start_hour, start_min, actual_duration)
        (0, 'Line A', 'SKU-1001 Granola Bar 40g', 'SKU-1002 Protein Bar 55g', 6, 0, 23),
        (0, 'Line A', 'SKU-1002 Protein Bar 55g', 'SKU-1003 Cereal Box 500g', 14, 30, 52),
        (0, 'Line B', 'SKU-1003 Cereal Box 500g', 'SKU-1007 Muesli 750g', 7, 15, 27),
        (0, 'Line B', 'SKU-1007 Muesli 750g', 'SKU-1004 Oat Crackers 200g', 13, 0, 58),
        (0, 'Line C', 'SKU-1005 Trail Mix 300g', 'SKU-1008 Nut Butter 250g', 8, 0, 38),
        (0, 'Line C', 'SKU-1008 Nut Butter 250g', 'SKU-1006 Rice Cakes 150g', 15, 45, 67),
        (0, 'Line D', 'SKU-1004 Oat Crackers 200g', 'SKU-1005 Trail Mix 300g', 6, 30, 49),
        (0, 'Line D', 'SKU-1006 Rice Cakes 150g', 'SKU-1001 Granola Bar 40g', 14, 0, 73),
        (1, 'Line A', 'SKU-1003 Cereal Box 500g', 'SKU-1007 Muesli 750g', 6, 0, 26),
        (1, 'Line A', 'SKU-1007 Muesli 750g', 'SKU-1005 Trail Mix 300g', 14, 15, 61),
        (1, 'Line B', 'SKU-1001 Granola Bar 40g', 'SKU-1002 Protein Bar 55g', 7, 30, 19),
        (1, 'Line B', 'SKU-1002 Protein Bar 55g', 'SKU-1008 Nut Butter 250g', 13, 45, 55),
        (1, 'Line C', 'SKU-1006 Rice Cakes 150g', 'SKU-1004 Oat Crackers 200g', 6, 0, 33),
        (1, 'Line C', 'SKU-1004 Oat Crackers 200g', 'SKU-1001 Granola Bar 40g', 14, 30, 48),
        (1, 'Line D', 'SKU-1008 Nut Butter 250g', 'SKU-1005 Trail Mix 300g', 8, 15, 37),
        (1, 'Line D', 'SKU-1005 Trail Mix 300g', 'SKU-1003 Cereal Box 500g', 15, 0, 79),
        (2, 'Line A', 'SKU-1002 Protein Bar 55g', 'SKU-1001 Granola Bar 40g', 6, 30, 18),
        (2, 'Line A', 'SKU-1001 Granola Bar 40g', 'SKU-1006 Rice Cakes 150g', 13, 45, 63),
        (2, 'Line B', 'SKU-1007 Muesli 750g', 'SKU-1003 Cereal Box 500g', 7, 0, 24),
        (2, 'Line B', 'SKU-1003 Cereal Box 500g', 'SKU-1004 Oat Crackers 200g', 14, 0, 57),
        (2, 'Line C', 'SKU-1008 Nut Butter 250g', 'SKU-1005 Trail Mix 300g', 6, 0, 34),
        (2, 'Line C', 'SKU-1005 Trail Mix 300g', 'SKU-1002 Protein Bar 55g', 15, 30, 72),
        (2, 'Line D', 'SKU-1001 Granola Bar 40g', 'SKU-1003 Cereal Box 500g', 7, 45, 51),
        (2, 'Line D', 'SKU-1006 Rice Cakes 150g', 'SKU-1008 Nut Butter 250g', 13, 30, 44),
        (3, 'Line A', 'SKU-1004 Oat Crackers 200g', 'SKU-1006 Rice Cakes 150g', 6, 0, 29),
        (3, 'Line A', 'SKU-1006 Rice Cakes 150g', 'SKU-1007 Muesli 750g', 14, 45, 68),
        (3, 'Line B', 'SKU-1005 Trail Mix 300g', 'SKU-1008 Nut Butter 250g', 7, 15, 36),
        (3, 'Line B', 'SKU-1008 Nut Butter 250g', 'SKU-1002 Protein Bar 55g', 13, 0, 82),
        (3, 'Line C', 'SKU-1003 Cereal Box 500g', 'SKU-1001 Granola Bar 40g', 6, 30, 47),
        (3, 'Line C', 'SKU-1001 Granola Bar 40g', 'SKU-1004 Oat Crackers 200g', 15, 15, 39),
        (3, 'Line D', 'SKU-1002 Protein Bar 55g', 'SKU-1007 Muesli 750g', 8, 0, 56),
        (3, 'Line D', 'SKU-1007 Muesli 750g', 'SKU-1006 Rice Cakes 150g', 14, 0, 43),
        (4, 'Line A', 'SKU-1008 Nut Butter 250g', 'SKU-1005 Trail Mix 300g', 6, 0, 33),
        (4, 'Line A', 'SKU-1005 Trail Mix 300g', 'SKU-1001 Granola Bar 40g', 13, 30, 78),
        (4, 'Line B', 'SKU-1006 Rice Cakes 150g', 'SKU-1004 Oat Crackers 200g', 7, 45, 31),
        (4, 'Line B', 'SKU-1004 Oat Crackers 200g', 'SKU-1003 Cereal Box 500g', 14, 30, 65),
        (4, 'Line C', 'SKU-1007 Muesli 750g', 'SKU-1003 Cereal Box 500g', 6, 15, 22),
        (4, 'Line C', 'SKU-1003 Cereal Box 500g', 'SKU-1002 Protein Bar 55g', 15, 0, 59),
        (4, 'Line D', 'SKU-1001 Granola Bar 40g', 'SKU-1008 Nut Butter 250g', 7, 30, 84),
        (4, 'Line D', 'SKU-1002 Protein Bar 55g', 'SKU-1006 Rice Cakes 150g', 13, 45, 46),
    ]

    for row_idx, rec in enumerate(records, 2):
        date_idx, line, from_p, to_p, start_h, start_m, duration = rec
        rec_date = week_dates[date_idx]
        start_dt = datetime(rec_date.year, rec_date.month, rec_date.day, start_h, start_m, 0)
        end_dt = start_dt + timedelta(minutes=duration)

        # Convert to time fractions (Excel serial time = fraction of day)
        start_time_val = start_dt.time()
        end_time_val = end_dt.time()

        # Target
        target = targets.get((from_p, to_p), default_target)

        # Units per hour for to_product (incoming product)
        units_ph = uph.get(to_p, 720)

        # Write columns A-F, H, K (leave G, I, J empty)
        ws1.cell(row=row_idx, column=1, value=rec_date)  # A: Date
        ws1.cell(row=row_idx, column=1).number_format = 'yyyy-mm-dd'
        ws1.cell(row=row_idx, column=2, value=line)       # B: Line
        ws1.cell(row=row_idx, column=3, value=from_p)     # C: From Product
        ws1.cell(row=row_idx, column=4, value=to_p)       # D: To Product
        ws1.cell(row=row_idx, column=5, value=start_time_val)  # E: Changeover Start
        ws1.cell(row=row_idx, column=5).number_format = 'hh:mm'
        ws1.cell(row=row_idx, column=6, value=end_time_val)    # F: Changeover End
        ws1.cell(row=row_idx, column=6).number_format = 'hh:mm'
        # G (col 7): empty — Duration Minutes
        # H: Target Minutes
        ws1.cell(row=row_idx, column=8, value=target)
        # I (col 9): empty — Over Target
        # J (col 10): empty — Lost Production Units
        ws1.cell(row=row_idx, column=11, value=units_ph)  # K: Units per Hour

    ws1.freeze_panes = 'A2'

    # ------------------------------------------------------------------ #
    # Sheet 2: ChangeoverSummary
    # ------------------------------------------------------------------ #
    ws2 = wb.create_sheet('ChangeoverSummary')

    summary_headers = ['Transition', 'Count', 'Avg Duration', 'Avg Target', 'Avg Overrun']
    for col, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFFFF')
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws2.row_dimensions[1].height = 25
    ws2.column_dimensions['A'].width = 45
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 14
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 12

    # 10 common transitions (from the records above) — only column A filled
    transitions = [
        'SKU-1001 Granola Bar 40g → SKU-1002 Protein Bar 55g',
        'SKU-1002 Protein Bar 55g → SKU-1001 Granola Bar 40g',
        'SKU-1003 Cereal Box 500g → SKU-1007 Muesli 750g',
        'SKU-1007 Muesli 750g → SKU-1003 Cereal Box 500g',
        'SKU-1004 Oat Crackers 200g → SKU-1006 Rice Cakes 150g',
        'SKU-1006 Rice Cakes 150g → SKU-1004 Oat Crackers 200g',
        'SKU-1005 Trail Mix 300g → SKU-1008 Nut Butter 250g',
        'SKU-1008 Nut Butter 250g → SKU-1005 Trail Mix 300g',
        'SKU-1001 Granola Bar 40g → SKU-1003 Cereal Box 500g',
        'SKU-1005 Trail Mix 300g → SKU-1001 Granola Bar 40g',
    ]

    for row_idx, trans in enumerate(transitions, 2):
        ws2.cell(row=row_idx, column=1, value=trans)  # A: Transition (only A filled)
        # B, C, D, E: empty

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  ChangeoverLog: 40 rows of changeover data (G, I, J empty)')
    print(f'  ChangeoverSummary: 10 transitions listed (Count/Avg cols empty)')


create_initial()
