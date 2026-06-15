"""
Initial Setup: Apply data validation dropdowns for deal stages and add conditional formatting to a pipeline tracker.
Task ID: calc_gpm_018
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_018'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pipeline'

    # --- Header Row (Row 1) ---
    headers = ['Deal Name', 'Company', 'Value', 'Stage', 'Probability', 'Expected Value']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF003366', end_color='FF003366', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Data Rows (A2:C8) ---
    # Only populate Deal Name, Company, Value columns
    # D (Stage), E (Probability), F (Expected Value) left empty for the task
    data = [
        ['Enterprise Suite', 'Acme Corp', 150000],
        ['Cloud Migration', 'TechStart', 85000],
        ['Security Audit', 'FinBank', 62000],
        ['Data Platform', 'HealthCo', 220000],
        ['Mobile App', 'RetailMax', 95000],
        ['API Integration', 'LogiTech', 45000],
        ['Support Plan', 'EduLearn', 35000],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Format Value column (C) as dollar
    for r in range(2, 9):
        ws.cell(row=r, column=3).number_format = '$#,##0'

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
