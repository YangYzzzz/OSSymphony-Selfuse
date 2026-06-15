"""
Reward Script: Apply data validation dropdowns for deal stages and add conditional formatting to a pipeline tracker.
Task ID: calc_gpm_018
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Stage values correctly populated in D2:D8
  Component 2 (0.20): Data validation dropdown on D2:D8 with correct stage list
  Component 3 (0.20): Probability formulas in E2:E8 mapping stage to probability
  Component 4 (0.15): Expected Value formulas in F2:F8 (=C*E pattern)
  Component 5 (0.20): Conditional formatting rules on D2:D8
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_018'

# Expected stage values per row
EXPECTED_STAGES = {
    2: 'Negotiation',
    3: 'Proposal',
    4: 'Qualified',
    5: 'Verbal Commit',
    6: 'Prospect',
    7: 'Closed Won',
    8: 'Negotiation',
}

# Valid stage names for data validation
VALID_STAGES = ['Prospect', 'Qualified', 'Proposal', 'Negotiation', 'Verbal Commit', 'Closed Won', 'Closed Lost']

# Stage to probability mapping
STAGE_PROB = {
    'Prospect': 0.1,
    'Qualified': 0.25,
    'Proposal': 0.5,
    'Negotiation': 0.7,
    'Verbal Commit': 0.9,
    'Closed Won': 1.0,
    'Closed Lost': 0.0,
}

# C column values (pre-existing)
C_VALUES = {2: 150000, 3: 85000, 4: 62000, 5: 220000, 6: 95000, 7: 45000, 8: 35000}


def persist_app_state(domain):
    """Attempt to save any unsaved GUI state via Ctrl+S."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Pipeline' not in wb.sheetnames:
        print("CRITICAL: 'Pipeline' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Pipeline']

    # =========================================================================
    # Component 1: Stage values correctly populated in D2:D8 (0.25 points)
    # Initial env has D2:D8 all None; golden has correct stage values.
    # =========================================================================
    try:
        correct_stages = 0
        for row in range(2, 9):
            cell_val = ws.cell(row=row, column=4).value  # column D
            expected = EXPECTED_STAGES[row]
            if cell_val is not None and str(cell_val).strip() == expected:
                correct_stages += 1
            else:
                print(f"  FAIL: D{row} expected '{expected}', found {cell_val!r}")

        if correct_stages == 7:
            print(f"PASS: Component 1 — All 7 stage values correct (0.25 pts)")
            total_score += 0.25
        elif correct_stages >= 4:
            partial = round(0.25 * (correct_stages / 7), 2)
            print(f"PARTIAL: Component 1 — {correct_stages}/7 stages correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {correct_stages}/7 stages correct")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Data validation dropdown on D2:D8 (0.20 points)
    # Initial env has 0 data validations; golden has 1 list validation.
    # =========================================================================
    try:
        validations = ws.data_validations.dataValidation
        list_dv_found = False
        for dv in validations:
            if dv.type == 'list':
                # Check if it covers D2:D8 range
                sqref_str = str(dv.sqref)
                # Check formula1 contains the required stages
                formula = dv.formula1 if dv.formula1 else ''
                has_stages = all(stage in formula for stage in VALID_STAGES)
                # Check the range covers D2:D8 (could be D2:D8 or similar)
                covers_d = 'D' in sqref_str.upper()

                if has_stages and covers_d:
                    print(f"PASS: Component 2 — Data validation list found on {sqref_str} with all stages (0.20 pts)")
                    total_score += 0.20
                    list_dv_found = True
                    break
                elif covers_d:
                    # Partial: has list DV on D but missing some stages
                    print(f"PARTIAL: Component 2 — Data validation on {sqref_str} but missing some stages (0.10 pts)")
                    total_score += 0.10
                    list_dv_found = True
                    break

        if not list_dv_found:
            print(f"FAIL: Component 2 — No list data validation found covering D2:D8")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Probability formulas in E2:E8 (0.20 points)
    # Initial env has E2:E8 all None; golden has nested IF formulas.
    # We check that E cells contain formulas (IF or VLOOKUP) that map stage to prob.
    # =========================================================================
    try:
        formula_count = 0
        for row in range(2, 9):
            e_val = ws.cell(row=row, column=5).value  # column E
            if e_val is not None and isinstance(e_val, str) and '=' in str(e_val):
                # Has a formula - check it references the stage column or contains IF/VLOOKUP
                formula_upper = str(e_val).upper()
                if 'IF' in formula_upper or 'VLOOKUP' in formula_upper:
                    formula_count += 1
                else:
                    print(f"  INFO: E{row} has formula but no IF/VLOOKUP: {e_val[:60]}")
            else:
                print(f"  FAIL: E{row} expected formula, found {e_val!r}")

        if formula_count == 7:
            print(f"PASS: Component 3 — All 7 probability formulas present (0.20 pts)")
            total_score += 0.20
        elif formula_count >= 4:
            partial = round(0.20 * (formula_count / 7), 2)
            print(f"PARTIAL: Component 3 — {formula_count}/7 probability formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {formula_count}/7 probability formulas found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: Expected Value formulas in F2:F8 (0.15 points)
    # Initial env has F2:F8 all None; golden has =C*E formulas.
    # =========================================================================
    try:
        f_formula_count = 0
        for row in range(2, 9):
            f_val = ws.cell(row=row, column=6).value  # column F
            if f_val is not None and isinstance(f_val, str) and '=' in str(f_val):
                formula_upper = str(f_val).upper()
                # Check it references C and E columns for this row
                c_ref = f'C{row}'
                e_ref = f'E{row}'
                if c_ref in formula_upper and e_ref in formula_upper:
                    f_formula_count += 1
                else:
                    # May use different reference style, still count if it's a formula
                    print(f"  INFO: F{row} has formula but doesn't reference C{row}*E{row}: {f_val[:60]}")
                    f_formula_count += 0.5
            else:
                print(f"  FAIL: F{row} expected formula, found {f_val!r}")

        if f_formula_count >= 7:
            print(f"PASS: Component 4 — All 7 expected value formulas correct (0.15 pts)")
            total_score += 0.15
        elif f_formula_count >= 4:
            partial = round(0.15 * (f_formula_count / 7), 2)
            print(f"PARTIAL: Component 4 — {f_formula_count}/7 expected value formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {f_formula_count}/7 expected value formulas found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # =========================================================================
    # Component 5: Conditional formatting rules on D2:D8 (0.20 points)
    # Initial env has 0 conditional formatting rules; golden has 4 rules.
    # =========================================================================
    try:
        cf_rules_on_d = 0
        cf_total_rules = 0
        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            # Check if the CF range includes column D rows 2-8
            if 'D' in cf_range.upper():
                for rule in cf.rules:
                    cf_total_rules += 1

        if cf_total_rules >= 4:
            print(f"PASS: Component 5 — {cf_total_rules} conditional formatting rules found on D range (0.20 pts)")
            total_score += 0.20
        elif cf_total_rules >= 2:
            partial = round(0.20 * (cf_total_rules / 4), 2)
            print(f"PARTIAL: Component 5 — {cf_total_rules}/4 conditional formatting rules ({partial} pts)")
            total_score += partial
        elif cf_total_rules >= 1:
            print(f"PARTIAL: Component 5 — Only {cf_total_rules} conditional formatting rule (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — No conditional formatting rules found on D2:D8")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
