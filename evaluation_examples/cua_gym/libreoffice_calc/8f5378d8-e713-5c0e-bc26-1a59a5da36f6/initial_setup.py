"""
Initial Setup: Pie chart with no data labels on Portfolio sheet
Task ID: calc_chart_data_labels_category_022
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.chart import PieChart, Reference

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_data_labels_category_022'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Portfolio ---
    ws = wb.active
    ws.title = 'Portfolio'

    # Headers
    ws['A1'] = 'Asset Class'
    ws['B1'] = 'Allocation %'

    # Data rows matching context
    data = [
        ('Stocks', 55),
        ('Bonds', 25),
        ('Real Estate', 12),
        ('Cash', 5),
        ('Commodities', 3),
    ]
    for r, (asset, alloc) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=asset)
        ws.cell(row=r, column=2, value=alloc)

    # --- Pie Chart (no data labels) ---
    pie = PieChart()
    pie.title = 'Portfolio Allocation'

    # Data reference: B1:B6 (includes header as title)
    data_ref = Reference(ws, min_col=2, min_row=1, max_row=6)
    pie.add_data(data_ref, titles_from_data=True)

    # Category reference: A2:A6 (asset class names)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=6)
    pie.set_categories(cats_ref)

    # NO data labels — that is what the task asks the agent to add
    # (pie.dataLabels is not set, so no labels appear)

    pie.width = 15
    pie.height = 12
    ws.add_chart(pie, 'D2')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
