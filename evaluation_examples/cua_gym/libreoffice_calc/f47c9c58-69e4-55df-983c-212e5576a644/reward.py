"""
Reward Script: Credit Limit Check - Available Credit, Utilization %, Conditional Formatting
Task ID: calc_fin_credit_limit_check_047
Domain: libreoffice_calc
Scoring:
  - Component 1: Headers D1/E1 added, row 1 bold, freeze panes A2  (0.15 pts)
  - Component 2: D2:D45 Available Credit formulas =Bn-Cn             (0.20 pts)
  - Component 3: E2:E45 Utilization % formulas =Cn/Bn + % format    (0.20 pts)
  - Component 4: Conditional formatting on E2:E45 (red>1, orange 0.8-1) (0.20 pts)
  - Component 5: Conditional formatting on D2:D45 (red font <0)     (0.10 pts)
  - Component 6: Summary rows 47-49 (COUNTIF, COUNTIFS, SUM, bold)  (0.15 pts)
  Total: 1.00
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_credit_limit_check_047'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: CreditMonitor sheet must exist
    if 'CreditMonitor' not in wb.sheetnames:
        print("CRITICAL: 'CreditMonitor' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['CreditMonitor']

    # Component 1: Headers D1/E1 added, row 1 bold, freeze panes A2 (0.15 pts)
    # These all FAIL on the initial file (D1/E1 are None, no freeze)
    try:
        d1 = ws.cell(1, 4).value
        e1 = ws.cell(1, 5).value
        a1_bold = ws.cell(1, 1).font.bold
        d1_bold = ws.cell(1, 4).font.bold
        freeze = ws.freeze_panes

        d1_ok = d1 is not None and str(d1).strip() == 'Available Credit'
        e1_ok = e1 is not None and str(e1).strip() == 'Utilization %'
        bold_ok = a1_bold == True and d1_bold == True
        freeze_ok = freeze == 'A2'

        if d1_ok and e1_ok and bold_ok and freeze_ok:
            print(f"PASS: Component 1 — Headers D1='{d1}', E1='{e1}', bold=True, freeze=A2 (0.15 pts)")
            total_score += 0.15
        else:
            details = []
            if not d1_ok:
                details.append(f"D1={repr(d1)} (expected 'Available Credit')")
            if not e1_ok:
                details.append(f"E1={repr(e1)} (expected 'Utilization %')")
            if not bold_ok:
                details.append(f"row1 bold: A1={a1_bold}, D1={d1_bold} (expected True)")
            if not freeze_ok:
                details.append(f"freeze_panes={repr(freeze)} (expected 'A2')")
            print(f"FAIL: Component 1 — {'; '.join(details)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: D2:D45 Available Credit formulas =Bn-Cn (0.20 pts)
    # Initial file has D2:D45 all None
    try:
        formula_count = 0
        bad_formulas = []
        for row in range(2, 46):
            val = ws.cell(row, 4).value
            expected = f'=B{row}-C{row}'
            if val is not None and isinstance(val, str):
                # Accept case-insensitive match
                if val.strip().upper() == expected.upper():
                    formula_count += 1
                else:
                    bad_formulas.append(f"D{row}={repr(val)} (expected {expected})")
            else:
                bad_formulas.append(f"D{row}={repr(val)} (expected {expected})")

        if formula_count == 44 and not bad_formulas:
            print(f"PASS: Component 2 — All 44 D2:D45 formulas correct (=Bn-Cn) (0.20 pts)")
            total_score += 0.20
        elif formula_count >= 40:
            print(f"PARTIAL: Component 2 — {formula_count}/44 D formulas correct (first issues: {bad_formulas[:3]})")
            # No partial credit for this component — requires all 44
            print(f"FAIL: Component 2 — need all 44, got {formula_count}")
        else:
            print(f"FAIL: Component 2 — only {formula_count}/44 D formulas correct. First issues: {bad_formulas[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: E2:E45 Utilization % formulas =Cn/Bn + percentage format (0.20 pts)
    # Initial file has E2:E45 all None
    try:
        formula_count_e = 0
        bad_formulas_e = []
        fmt_count = 0
        for row in range(2, 46):
            val = ws.cell(row, 5).value
            expected = f'=C{row}/B{row}'
            cell_fmt = ws.cell(row, 5).number_format
            if val is not None and isinstance(val, str):
                if val.strip().upper() == expected.upper():
                    formula_count_e += 1
                else:
                    bad_formulas_e.append(f"E{row}={repr(val)}")
            else:
                bad_formulas_e.append(f"E{row}={repr(val)}")
            # Check percentage format
            if cell_fmt and '%' in str(cell_fmt):
                fmt_count += 1

        formulas_ok = (formula_count_e == 44)
        fmt_ok = (fmt_count >= 40)  # allow a few without format (format sometimes inherited)

        if formulas_ok and fmt_ok:
            print(f"PASS: Component 3 — All 44 E2:E45 formulas correct (=Cn/Bn) + percentage format ({fmt_count}/44 formatted) (0.20 pts)")
            total_score += 0.20
        else:
            details = []
            if not formulas_ok:
                details.append(f"{formula_count_e}/44 formulas correct (first issues: {bad_formulas_e[:3]})")
            if not fmt_ok:
                details.append(f"{fmt_count}/44 cells have percentage format")
            print(f"FAIL: Component 3 — {'; '.join(details)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Conditional formatting on E2:E45 (red >1, orange 0.8-1) (0.20 pts)
    # Initial file has no conditional formatting
    try:
        e_cf_ranges = []
        for cf_obj in ws.conditional_formatting:
            cf_str = str(cf_obj)
            if 'E2' in cf_str and 'E45' in cf_str:
                e_cf_ranges.append(cf_obj)

        if not e_cf_ranges:
            print("FAIL: Component 4 — No conditional formatting found for E2:E45")
        else:
            # Get all rules for E2:E45 range
            rules_list = []
            for cf_obj in e_cf_ranges:
                rules_list.extend(ws.conditional_formatting._cf_rules[cf_obj])

            red_rule_found = False
            orange_rule_found = False

            for rule in rules_list:
                formula = rule.formula if hasattr(rule, 'formula') else []
                op = rule.operator if hasattr(rule, 'operator') else None
                dxf = rule.dxf

                if dxf and dxf.fill:
                    try:
                        fill_color = dxf.fill.fgColor.rgb
                    except:
                        fill_color = None

                    # Red fill (FFFF0000) for >1 (greaterThan) or >100%
                    if fill_color and 'FF0000' in fill_color:
                        if op == 'greaterThan' and formula and formula[0] in ['1', '1.0', '100%', '1.00']:
                            red_rule_found = True

                    # Orange fill for between 0.8 and 1
                    if fill_color and 'FFA500' in fill_color:
                        if op == 'between' and len(formula) >= 2:
                            lo = formula[0]
                            hi = formula[1]
                            if lo in ['0.8', '0.80', '80%'] and hi in ['1', '1.0', '100%']:
                                orange_rule_found = True

            if red_rule_found and orange_rule_found:
                print("PASS: Component 4 — E2:E45 CF: red for >1 and orange for 0.8-1 found (0.20 pts)")
                total_score += 0.20
            else:
                details = []
                if not red_rule_found:
                    details.append("missing red (>1) rule")
                if not orange_rule_found:
                    details.append("missing orange (0.8-1) rule")
                # Show all rules for debugging
                rule_info = []
                for rule in rules_list:
                    op2 = getattr(rule, 'operator', None)
                    frm = getattr(rule, 'formula', [])
                    clr = None
                    if rule.dxf and rule.dxf.fill:
                        try:
                            clr = rule.dxf.fill.fgColor.rgb
                        except:
                            pass
                    rule_info.append(f"op={op2},formula={frm},fill={clr}")
                print(f"FAIL: Component 4 — {'; '.join(details)}. Rules found: {rule_info}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Conditional formatting on D2:D45 (red font for negative/< 0) (0.10 pts)
    # Initial file has no conditional formatting
    try:
        d_cf_ranges = []
        for cf_obj in ws.conditional_formatting:
            cf_str = str(cf_obj)
            if 'D2' in cf_str and 'D45' in cf_str:
                d_cf_ranges.append(cf_obj)

        if not d_cf_ranges:
            print("FAIL: Component 5 — No conditional formatting found for D2:D45")
        else:
            rules_d = []
            for cf_obj in d_cf_ranges:
                rules_d.extend(ws.conditional_formatting._cf_rules[cf_obj])

            red_font_found = False
            for rule in rules_d:
                formula = rule.formula if hasattr(rule, 'formula') else []
                op = rule.operator if hasattr(rule, 'operator') else None
                dxf = rule.dxf

                if dxf and dxf.font:
                    try:
                        font_color = dxf.font.color.rgb
                    except:
                        font_color = None

                    # Red font (FFFF0000) for < 0 (lessThan)
                    if font_color and 'FF0000' in font_color:
                        if op == 'lessThan' and formula and formula[0] in ['0', '0.0']:
                            red_font_found = True

            if red_font_found:
                print("PASS: Component 5 — D2:D45 CF: red font for negative values found (0.10 pts)")
                total_score += 0.10
            else:
                rule_info = []
                for rule in rules_d:
                    op2 = getattr(rule, 'operator', None)
                    frm = getattr(rule, 'formula', [])
                    fclr = None
                    if rule.dxf and rule.dxf.font:
                        try:
                            fclr = rule.dxf.font.color.rgb
                        except:
                            pass
                    rule_info.append(f"op={op2},formula={frm},font={fclr}")
                print(f"FAIL: Component 5 — red font <0 rule not found. Rules: {rule_info}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Summary rows 47-49 (COUNTIF, COUNTIFS, SUM + bold C49) (0.15 pts)
    # Initial file has rows 47-49 all empty
    try:
        a47 = ws.cell(47, 1).value
        b47 = ws.cell(47, 2).value
        a48 = ws.cell(48, 1).value
        b48 = ws.cell(48, 2).value
        a49 = ws.cell(49, 1).value
        c49 = ws.cell(49, 3).value
        c49_bold = ws.cell(49, 3).font.bold
        a49_bold = ws.cell(49, 1).font.bold

        a47_ok = a47 is not None and 'Over Limit' in str(a47)
        b47_ok = b47 is not None and isinstance(b47, str) and 'COUNTIF' in b47.upper() and 'E2:E45' in b47.upper()
        a48_ok = a48 is not None and ('High Utilization' in str(a48) or 'Utilization' in str(a48))
        b48_ok = b48 is not None and isinstance(b48, str) and 'COUNTIFS' in b48.upper() and 'E2:E45' in b48.upper()
        a49_ok = a49 is not None and 'Total' in str(a49) and 'Outstanding' in str(a49)
        c49_ok = c49 is not None and isinstance(c49, str) and 'SUM' in c49.upper() and 'C2:C45' in c49.upper()
        bold_ok_49 = c49_bold == True or a49_bold == True

        all_ok = a47_ok and b47_ok and a48_ok and b48_ok and a49_ok and c49_ok and bold_ok_49

        if all_ok:
            print(f"PASS: Component 6 — Summary rows 47-49: COUNTIF, COUNTIFS, SUM formulas + bold (0.15 pts)")
            total_score += 0.15
        else:
            details = []
            if not a47_ok: details.append(f"A47={repr(a47)} (expected 'Over Limit Count')")
            if not b47_ok: details.append(f"B47={repr(b47)} (expected COUNTIF formula)")
            if not a48_ok: details.append(f"A48={repr(a48)} (expected 'High Utilization Count')")
            if not b48_ok: details.append(f"B48={repr(b48)} (expected COUNTIFS formula)")
            if not a49_ok: details.append(f"A49={repr(a49)} (expected 'Total Outstanding')")
            if not c49_ok: details.append(f"C49={repr(c49)} (expected SUM formula)")
            if not bold_ok_49: details.append(f"C49 bold={c49_bold}, A49 bold={a49_bold} (expected True)")
            print(f"FAIL: Component 6 — {'; '.join(details)}")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
