"""
Initial Setup: Credit Limit Check - CreditMonitor spreadsheet
Task ID: calc_fin_credit_limit_check_047
Domain: libreoffice_calc

Creates CreditMonitor sheet with customers, credit limits, and outstanding balances.
Columns D and E are intentionally left empty (to be filled by the agent task).
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_credit_limit_check_047'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: CreditMonitor ---
    ws = wb.active
    ws.title = 'CreditMonitor'

    # Headers (Row 1) - bold
    headers = ['Customer', 'Credit Limit', 'Outstanding Balance']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Column widths for readability
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 22

    # Realistic customer data - 44 rows (rows 2-45)
    customers = [
        ('Blackwell Industries',        85000,   72300),
        ('Horizon Tech Solutions',      120000,  115800),
        ('Coastal Marine Supplies',      45000,   12500),
        ('Summit Financial Group',      200000,  198700),
        ('Prairie Wind Energy',          60000,   48200),
        ('Northgate Retail Corp',       150000,   89500),
        ('Evergreen Landscaping',        30000,   31200),  # over limit
        ('Atlas Logistics Ltd',          75000,   61000),
        ('Meridian Healthcare',         100000,   93400),
        ('Redwood Construction',         55000,   44800),
        ('Cascade Electronics',          90000,   76500),
        ('Phoenix Distribution',         40000,   10800),
        ('Blue Ridge Consulting',        70000,   69200),
        ('Ironclad Security',            35000,   28700),
        ('Pacific Rim Imports',         110000,  110500),  # over limit
        ('Maple Leaf Properties',        80000,   22100),
        ('Thunder Bay Manufacturing',    95000,   86400),
        ('Silver Creek Winery',          25000,   24800),
        ('Cornerstone Capital',         180000,  162000),
        ('Desert Sun Agriculture',       42000,   38900),
        ('Lakeview Medical Group',       65000,   53200),
        ('Granite Peak Mining',         130000,   78600),
        ('Skyline Architecture',         50000,   45100),
        ('Amber Valley Foods',           38000,   37500),
        ('Centurion Auto Group',         88000,   71200),
        ('Willow Creek Nursery',         22000,    8900),
        ('Broadfield Insurance',         75000,   60500),
        ('Aurora Pharmaceuticals',      140000,  139800),  # almost over
        ('Titan Industrial',             58000,   52300),
        ('Clearwater Shipping',          82000,   34600),
        ('Highland Textile',             47000,   46800),
        ('Metro Catering Services',      31000,   32500),  # over limit
        ('Olympus Publishing',           68000,   57900),
        ('Crimson Bay Resorts',          95000,   85600),
        ('Greenfield Organics',          28000,   16400),
        ('Nexus Technology',            160000,  128000),
        ('Falcon Creek Brewing',         35000,   33800),
        ('Stonehaven Law Firm',          72000,   70100),
        ('Bayside Dental Group',         45000,   18200),
        ('Westport Trading Co',         105000,   84200),
        ('Riverbend Furniture',          62000,   58900),
        ('Northstar Aviation',          175000,  171500),  # over limit
        ('Coral Springs Realty',         53000,   42700),
        ('Pinnacle Staffing',            40000,   35900),
    ]

    for r, (customer, limit, balance) in enumerate(customers, 2):
        ws.cell(row=r, column=1, value=customer)
        ws.cell(row=r, column=2, value=limit)
        ws.cell(row=r, column=3, value=balance)

    # NOTE: Columns D and E are intentionally EMPTY
    # The task requires the agent to add:
    #   D1: 'Available Credit', E1: 'Utilization %'
    #   D2:D45 formulas, E2:E45 formulas
    #   Conditional formatting, summary rows 47-49

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: CreditMonitor')
    print(f'  Rows: 1 header + 44 data rows (rows 2-45)')
    print(f'  Columns: A=Customer, B=Credit Limit, C=Outstanding Balance')
    print(f'  Columns D, E: empty (for agent to fill)')


create_initial()
