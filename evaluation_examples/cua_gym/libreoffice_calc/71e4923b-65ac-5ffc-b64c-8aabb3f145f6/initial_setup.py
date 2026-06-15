"""
Initial Setup: Employee Attendance Sheet (No Conditional Formatting)
Task ID: osworld_calc_conditional_format_weekday_003
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_conditional_format_weekday_003'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Attendance ---
    ws = wb.active
    ws.title = "Attendance"

    # Headers
    headers = ["Attendance Date", "Employee Name", "Status", "Hours"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10

    # Realistic employee names
    employees = [
        "Sarah Chen",
        "Marcus Johnson",
        "Priya Patel",
        "Derek Williams",
        "Amina Hassan",
        "Lucas Fernandez",
        "Chloe Thompson",
        "Noah Kim",
        "Elena Reyes",
        "Benjamin Carter",
        "Fatima Al-Rashidi",
        "Tyler Brooks",
        "Mei-Ling Zhou",
        "Andre Dubois",
        "Ingrid Larsen",
    ]

    statuses = ["Present", "Absent", "Late", "Half-Day", "WFH"]
    status_weights = [0.65, 0.08, 0.10, 0.07, 0.10]

    # Generate ~100 rows of attendance data starting from 2025-01-06 (Monday)
    import random
    random.seed(42)

    start_date = date(2025, 1, 6)
    row = 2
    current_date = start_date

    # Build 100 rows: iterate through dates, assign employees
    generated = 0
    emp_idx = 0
    while generated < 100:
        emp_name = employees[emp_idx % len(employees)]
        emp_idx += 1

        # Pick status
        r = random.random()
        cumulative = 0
        chosen_status = "Present"
        for s, w in zip(statuses, status_weights):
            cumulative += w
            if r <= cumulative:
                chosen_status = s
                break

        # Hours based on status
        if chosen_status == "Present":
            hours = round(random.uniform(7.5, 9.0), 1)
        elif chosen_status == "Absent":
            hours = 0
        elif chosen_status == "Late":
            hours = round(random.uniform(5.0, 7.5), 1)
        elif chosen_status == "Half-Day":
            hours = round(random.uniform(3.5, 4.5), 1)
        else:  # WFH
            hours = round(random.uniform(7.0, 8.5), 1)

        ws.cell(row=row, column=1, value=current_date)
        ws.cell(row=row, column=1).number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=2, value=emp_name)
        ws.cell(row=row, column=3, value=chosen_status)
        ws.cell(row=row, column=4, value=hours)

        row += 1
        generated += 1

        # Advance date every 15 employees (simulate ~2 weeks of data per block)
        if emp_idx % len(employees) == 0:
            current_date += timedelta(days=1)
            # Skip nothing — include weekends intentionally so task makes sense

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
