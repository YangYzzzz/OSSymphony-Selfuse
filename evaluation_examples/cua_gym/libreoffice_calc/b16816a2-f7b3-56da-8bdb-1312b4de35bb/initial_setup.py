"""
Initial Setup: Terminal + LibreOffice Calc multi-app task
Task ID: osworld_multi_apps_terminal_calc_011
Domain: libreoffice_calc (multi-app: terminal + calc)

Creates:
  - /home/user/Desktop/item_ids.xlsx     (one column: ItemID, 12 data rows)
  - /home/user/Desktop/warehouse_quantities.ods  (one column: Quantity, 12 data rows)
Opens:
  - A terminal window so the agent can use the command line
"""

import os
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user/Desktop'
TASK_ID = 'osworld_multi_apps_terminal_calc_011'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_item_ids_xlsx():
    """Create item_ids.xlsx with a single ItemID column on the Desktop."""
    output_path = f'{WORKDIR}/item_ids.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ItemIDs'

    # Header
    ws.cell(row=1, column=1, value='ItemID')

    # Realistic item IDs (SKU-style codes)
    item_ids = [
        'SKU-1001',
        'SKU-1002',
        'SKU-1003',
        'SKU-1004',
        'SKU-1005',
        'SKU-1006',
        'SKU-1007',
        'SKU-1008',
        'SKU-1009',
        'SKU-1010',
        'SKU-1011',
        'SKU-1012',
    ]

    for r, item_id in enumerate(item_ids, 2):
        ws.cell(row=r, column=1, value=item_id)

    wb.save(output_path)
    print(f'Created: {output_path}')
    return item_ids


def create_warehouse_quantities_ods(item_ids):
    """
    Create warehouse_quantities.ods with a single Quantity column.
    Uses openpyxl to create an xlsx first, then convert to ods via LibreOffice,
    OR write directly as a CSV-embedded ODS using a simple approach.
    Since openpyxl doesn't support .ods natively, we use the odfpy approach.
    """
    output_path = f'{WORKDIR}/warehouse_quantities.ods'

    # Quantities: mix of low-stock (< 5, will be "Reorder") and normal stock
    # Align with item_ids list (12 items)
    quantities = [
        23,   # SKU-1001 - OK
        3,    # SKU-1002 - Reorder
        47,   # SKU-1003 - OK
        2,    # SKU-1004 - Reorder
        15,   # SKU-1005 - OK
        1,    # SKU-1006 - Reorder
        88,   # SKU-1007 - OK
        4,    # SKU-1008 - Reorder
        31,   # SKU-1009 - OK
        0,    # SKU-1010 - Reorder
        56,   # SKU-1011 - OK
        12,   # SKU-1012 - OK
    ]

    # Write as ODS using ezodf or odfpy if available; fallback to writing CSV and converting
    # We'll try odfpy first, then fallback to LibreOffice conversion
    try:
        import odf.opendocument as opendocument
        import odf.table as odftable
        import odf.text as odftext
        from odf.opendocument import OpenDocumentSpreadsheet
        from odf.table import Table, TableRow, TableCell
        from odf.text import P

        doc = OpenDocumentSpreadsheet()
        table = Table(name='Quantities')
        doc.spreadsheet.addElement(table)

        # Header row
        tr = TableRow()
        table.addElement(tr)
        tc = TableCell(valuetype='string')
        tc.addElement(P(text='Quantity'))
        tr.addElement(tc)

        # Data rows
        for qty in quantities:
            tr = TableRow()
            table.addElement(tr)
            tc = TableCell(valuetype='float', value=str(qty))
            tc.addElement(P(text=str(qty)))
            tr.addElement(tc)

        doc.save(output_path)
        print(f'Created (odfpy): {output_path}')

    except ImportError:
        # Fallback: create xlsx, then convert using LibreOffice headless
        tmp_xlsx = f'{WORKDIR}/warehouse_quantities_tmp.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Quantities'
        ws.cell(row=1, column=1, value='Quantity')
        for r, qty in enumerate(quantities, 2):
            ws.cell(row=r, column=1, value=qty)
        wb.save(tmp_xlsx)

        # Convert to ODS using LibreOffice headless
        env = os.environ.copy()
        env["DISPLAY"] = ":0"
        result = subprocess.run(
            ['libreoffice', '--headless', '--convert-to', 'ods',
             '--outdir', WORKDIR, tmp_xlsx],
            capture_output=True, text=True, env=env, timeout=30
        )
        print(f'LibreOffice conversion stdout: {result.stdout}')
        print(f'LibreOffice conversion stderr: {result.stderr}')

        # Rename if needed
        converted = f'{WORKDIR}/warehouse_quantities_tmp.ods'
        if os.path.exists(converted):
            os.rename(converted, output_path)
            print(f'Renamed to: {output_path}')

        # Cleanup temp file
        if os.path.exists(tmp_xlsx):
            os.remove(tmp_xlsx)

        if os.path.exists(output_path):
            print(f'Created (via LO conversion): {output_path}')
        else:
            print(f'ERROR: Failed to create {output_path}')

    return quantities


def create_initial():
    os.makedirs(WORKDIR, exist_ok=True)

    # Make sure inventory_check.csv does NOT exist in initial state
    csv_path = f'{WORKDIR}/inventory_check.csv'
    if os.path.exists(csv_path):
        os.remove(csv_path)
        print(f'Removed pre-existing: {csv_path}')

    # Create the two source files
    item_ids = create_item_ids_xlsx()
    create_warehouse_quantities_ods(item_ids)

    # GUI-ready startup: open a terminal
    # gnome-terminal is typically available on OSWorld Ubuntu VMs
    launch_gui('gnome-terminal', delay_sec=2.0)
    print('GUI_READY: launched terminal with DISPLAY=:0')


create_initial()
