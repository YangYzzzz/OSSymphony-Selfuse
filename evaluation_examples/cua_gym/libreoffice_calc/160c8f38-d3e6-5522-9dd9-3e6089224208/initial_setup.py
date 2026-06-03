"""
Initial Setup: Daily Sales spreadsheet with revenue data, no conditional formatting
Task ID: calc_gcv_032
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_032'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def get_business_days(start, count):
    """Generate a list of business days (Mon-Fri)."""
    days = []
    current = start
    while len(days) < count:
        if current.weekday() < 5:  # Mon=0 .. Fri=4
            days.append(current)
        current += timedelta(days=1)
    return days

def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Daily_Sales'

    # --- Headers ---
    headers = ['Date', 'Store', 'Daily Revenue']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Data: 39 rows of daily sales ---
    stores = [
        'Downtown Flagship', 'Westside Mall', 'Harbor View',
        'Northgate Plaza', 'East Village', 'Riverside Market',
        'Sunset Boulevard', 'Central Station', 'Lakewood Center',
        'Park Avenue'
    ]

    business_days = get_business_days(date(2025, 3, 3), 39)

    # Generate revenue data with some >10% jumps baked in
    revenues = []
    prev = random.randint(12000, 18000)
    revenues.append(prev)
    for i in range(1, 39):
        # Occasionally create a >10% jump (about 30% of the time)
        if random.random() < 0.3:
            change = random.uniform(0.12, 0.35)
            val = int(prev * (1 + change))
        else:
            change = random.uniform(-0.08, 0.08)
            val = int(prev * (1 + change))
        val = max(5000, min(35000, val))
        revenues.append(val)
        prev = val

    currency_format = '$#,##0'
    date_format = 'yyyy-mm-dd'

    for i in range(39):
        row = i + 2
        # Date
        cell_date = ws.cell(row=row, column=1, value=business_days[i])
        cell_date.number_format = date_format
        cell_date.border = thin_border

        # Store (cycle through stores)
        cell_store = ws.cell(row=row, column=2, value=stores[i % len(stores)])
        cell_store.border = thin_border

        # Daily Revenue
        cell_rev = ws.cell(row=row, column=3, value=revenues[i])
        cell_rev.number_format = currency_format
        cell_rev.border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 16

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
