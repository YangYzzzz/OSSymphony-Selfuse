"""
Reward Script: Build diversity metrics dashboard
Task ID: calc_hr_workforce_diversity_070
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Column C (C2:C14) has =Bn/$B$15 formulas + 0.0% format   — 0.30 pts
  Component 2: Column E (E2:E14) has =Cn-Dn formulas + 0.0% format       — 0.25 pts
  Component 3: Conditional formatting on E2:E14 (green E>0, red E<0)     — 0.20 pts
  Component 4: Grouped bar chart with correct title on the sheet          — 0.15 pts
  Component 5: Chart has 2 series referencing ethnicity rows (C9:C14, D9:D14) — 0.10 pts
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_workforce_diversity_070'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: sheet must exist
    if 'Diversity Data' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Diversity Data' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Diversity Data']

    # -----------------------------------------------------------------------
    # Component 1: Column C formulas =Bn/$B$15 with 0.0% format (0.30 pts)
    # For rows 2-14: C cell must contain formula like =B2/$B$15 and be formatted as 0.0%
    # In initial file: C2:C14 are all None — so this FAILS on initial
    # -----------------------------------------------------------------------
    try:
        c_formula_count = 0
        c_format_count = 0
        expected_formula_pattern = re.compile(r'^=B(\d+)/\$B\$15$', re.IGNORECASE)

        for row in range(2, 15):
            cell = ws.cell(row=row, column=3)
            val = cell.value
            fmt = cell.number_format

            if val and isinstance(val, str):
                match = expected_formula_pattern.match(val.replace(' ', ''))
                if match and int(match.group(1)) == row:
                    c_formula_count += 1
                else:
                    print(f"FAIL: C{row} formula mismatch: {repr(val)}")
            else:
                print(f"FAIL: C{row} has no formula: {repr(val)}")

            if fmt == '0.0%':
                c_format_count += 1
            else:
                print(f"FAIL: C{row} number_format mismatch: {repr(fmt)}, expected '0.0%'")

        # Award points only if all 13 cells have correct formulas AND correct format
        if c_formula_count == 13 and c_format_count == 13:
            print(f"PASS: Component 1 — All 13 C column formulas correct with 0.0% format (0.30 pts)")
            total_score += 0.30
        elif c_formula_count == 13:
            print(f"PARTIAL: Component 1 — Formulas correct but format wrong ({c_format_count}/13 formatted) (0.15 pts)")
            total_score += 0.15
        elif c_formula_count > 0:
            print(f"PARTIAL: Component 1 — Only {c_formula_count}/13 C formulas correct (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 — No C column formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Column E formulas =Cn-Dn with 0.0% format (0.25 pts)
    # For rows 2-14: E cell must contain formula like =C2-D2 and be formatted as 0.0%
    # In initial file: E2:E14 are all None — so this FAILS on initial
    # -----------------------------------------------------------------------
    try:
        e_formula_count = 0
        e_format_count = 0
        expected_e_pattern = re.compile(r'^=C(\d+)-D(\d+)$', re.IGNORECASE)

        for row in range(2, 15):
            cell = ws.cell(row=row, column=5)
            val = cell.value
            fmt = cell.number_format

            if val and isinstance(val, str):
                match = expected_e_pattern.match(val.replace(' ', ''))
                if match and int(match.group(1)) == row and int(match.group(2)) == row:
                    e_formula_count += 1
                else:
                    print(f"FAIL: E{row} formula mismatch: {repr(val)}")
            else:
                print(f"FAIL: E{row} has no formula: {repr(val)}")

            if fmt == '0.0%':
                e_format_count += 1
            else:
                print(f"FAIL: E{row} number_format mismatch: {repr(fmt)}, expected '0.0%'")

        if e_formula_count == 13 and e_format_count == 13:
            print(f"PASS: Component 2 — All 13 E column formulas correct with 0.0% format (0.25 pts)")
            total_score += 0.25
        elif e_formula_count == 13:
            print(f"PARTIAL: Component 2 — Formulas correct but format wrong ({e_format_count}/13 formatted) (0.12 pts)")
            total_score += 0.12
        elif e_formula_count > 0:
            print(f"PARTIAL: Component 2 — Only {e_formula_count}/13 E formulas correct (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 2 — No E column formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Conditional formatting on E2:E14 — green font for E>0, red for E<0 (0.20 pts)
    # In initial file: no CF rules exist — FAILS on initial
    # Expected: two FormulaRule entries on E2:E14
    #   Rule 1: formula=['E2>0'], dxf.font.color = #70AD47 green
    #   Rule 2: formula=['E2<0'], dxf.font.color = #FF0000 red
    # -----------------------------------------------------------------------
    try:
        cf_rules = ws.conditional_formatting
        found_green_rule = False
        found_red_rule = False

        for cf_range_obj in cf_rules:
            cf_range_str = str(cf_range_obj)
            # Check if this CF applies to E2:E14 or overlapping range
            if 'E' in cf_range_str.upper():
                for rule in cf_range_obj.rules:
                    if rule.type == 'expression' and rule.formula:
                        formula_str = rule.formula[0].upper().replace(' ', '')
                        dxf_font_color = None
                        if rule.dxf and rule.dxf.font and rule.dxf.font.color:
                            try:
                                dxf_font_color = rule.dxf.font.color.rgb.upper()
                            except Exception:
                                pass

                        # Green rule: E2>0, color contains 70AD47
                        if 'E2>0' in formula_str or 'E>0' in formula_str:
                            if dxf_font_color and ('70AD47' in dxf_font_color):
                                found_green_rule = True
                                print(f"PASS: CF green rule found (E>0, color={dxf_font_color})")
                            else:
                                print(f"PARTIAL: CF green rule formula found but color wrong: {dxf_font_color}")

                        # Red rule: E2<0, color contains FF0000
                        if 'E2<0' in formula_str or 'E<0' in formula_str:
                            if dxf_font_color and ('FF0000' in dxf_font_color):
                                found_red_rule = True
                                print(f"PASS: CF red rule found (E<0, color={dxf_font_color})")
                            else:
                                print(f"PARTIAL: CF red rule formula found but color wrong: {dxf_font_color}")

        if found_green_rule and found_red_rule:
            print(f"PASS: Component 3 — Both conditional formatting rules correct (0.20 pts)")
            total_score += 0.20
        elif found_green_rule or found_red_rule:
            print(f"PARTIAL: Component 3 — Only one CF rule found (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — No conditional formatting rules found on column E")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: Grouped bar chart with title 'Ethnicity Representation: This Year vs Last Year' (0.15 pts)
    # In initial file: no charts exist — FAILS on initial
    # -----------------------------------------------------------------------
    try:
        charts = ws._charts
        if len(charts) >= 1:
            chart = charts[0]
            # Verify it is a BarChart
            is_bar_chart = chart.__class__.__name__ == 'BarChart'

            # Get chart title text
            chart_title_text = None
            try:
                chart_title_text = chart.title.tx.rich.p[0].r[0].t
            except Exception:
                pass

            expected_title = 'Ethnicity Representation: This Year vs Last Year'
            title_correct = (chart_title_text and
                             chart_title_text.strip() == expected_title)

            # Check grouping (clustered = grouped)
            grouping_ok = getattr(chart, 'grouping', None) in ('clustered', 'standard', None)

            if is_bar_chart and title_correct:
                print(f"PASS: Component 4 — BarChart with correct title found (0.15 pts)")
                total_score += 0.15
            elif is_bar_chart:
                print(f"PARTIAL: Component 4 — BarChart found but title wrong: {repr(chart_title_text)} (0.07 pts)")
                total_score += 0.07
            else:
                print(f"PARTIAL: Component 4 — Chart found but not BarChart: {chart.__class__.__name__} (0.05 pts)")
                total_score += 0.05
        else:
            print(f"FAIL: Component 4 — No charts found on sheet (expected 1)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -----------------------------------------------------------------------
    # Component 5: Chart has 2 series covering ethnicity rows (C9:C14 and D9:D14) (0.10 pts)
    # In initial file: no charts exist — FAILS on initial
    # -----------------------------------------------------------------------
    try:
        charts = ws._charts
        if len(charts) >= 1:
            chart = charts[0]
            num_series = len(chart.series)

            # Verify 2 series exist
            if num_series >= 2:
                # Check series reference rows — should be rows 9-14 (ethnicity data)
                ser0_ref = None
                ser1_ref = None
                try:
                    ser0_ref = chart.series[0].val.numRef.ref
                except Exception:
                    pass
                try:
                    ser1_ref = chart.series[1].val.numRef.ref
                except Exception:
                    pass

                # Series should reference rows 9-14 (ethnicity groups)
                # Expected: '$C$9:$C$14' and '$D$9:$D$14' (or similar row ranges including 9-14)
                def refs_ethnicity_rows(ref_str):
                    if ref_str is None:
                        return False
                    # Look for rows 9 through 14 in the reference
                    return bool(re.search(r'\$[CD]\$9:\$[CD]\$14', str(ref_str), re.IGNORECASE))

                ser0_ok = refs_ethnicity_rows(ser0_ref)
                ser1_ok = refs_ethnicity_rows(ser1_ref)

                # Also accept that series titles indicate This Year % and Last Year %
                ser0_title = None
                ser1_title = None
                try:
                    ser0_title = chart.series[0].title.v
                except Exception:
                    pass
                try:
                    ser1_title = chart.series[1].title.v
                except Exception:
                    pass

                if ser0_ok and ser1_ok:
                    print(f"PASS: Component 5 — 2 series covering ethnicity rows (C9:C14, D9:D14) (0.10 pts)")
                    total_score += 0.10
                elif num_series >= 2 and (ser0_ref or ser1_ref):
                    # 2 series exist but range might be slightly different — partial credit
                    print(f"PARTIAL: Component 5 — 2 series found but refs may not match ethnicity rows exactly: {ser0_ref}, {ser1_ref} (0.05 pts)")
                    total_score += 0.05
                else:
                    print(f"FAIL: Component 5 — 2 series found but refs missing: ser0={ser0_ref}, ser1={ser1_ref}")
            else:
                print(f"FAIL: Component 5 — Chart has {num_series} series, expected 2")
        else:
            print(f"FAIL: Component 5 — No charts found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
