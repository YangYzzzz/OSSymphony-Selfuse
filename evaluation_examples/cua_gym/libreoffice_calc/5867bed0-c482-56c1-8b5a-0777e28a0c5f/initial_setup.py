"""
Initial Setup: HR Workforce Diversity Metrics Dashboard
Task ID: calc_hr_workforce_diversity_070
Domain: libreoffice_calc

Creates the initial state of the diversity dashboard spreadsheet.
The sheet has category names, headcount (B), prior year % (D),
but C (This Year %) and E (Change) are intentionally left EMPTY
because the task requires the agent to fill them with formulas.
No chart exists in the initial state.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_workforce_diversity_070'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Diversity Data ---
    ws = wb.active
    ws.title = 'Diversity Data'

    # ---- Header row ----
    header_font = Font(name='Calibri', bold=True, size=11)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_font_white = Font(name='Calibri', bold=True, size=11, color='FFFFFFFF')

    headers = ['Category', 'This Year Count', 'This Year %', 'Last Year %', 'Change']
    col_widths = [28, 18, 15, 15, 12]
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 28

    # ---- Gender section label ----
    section_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    section_font = Font(name='Calibri', bold=True, size=11, italic=True)

    # Rows 2-8: Gender breakdown
    # Gender groups — realistic 2025 headcount data
    # Total = 348 employees
    # Male: 181, Female: 152, Non-Binary: 9, Prefer Not to Say: 6
    # Sum: 181+152+9+6 = 348  ✓
    gender_data = [
        # (Category, This Year Count, Last Year %)
        ('Male',                181,  0.5316),
        ('Female',              152,  0.4253),
        ('Non-Binary',            9,  0.0202),
        ('Prefer Not to Say',     6,  0.0144),
    ]

    # Rows 9-14: Ethnicity breakdown
    # White: 148, Hispanic/Latino: 72, Black/African American: 61,
    # Asian: 43, Two or More Races: 14, Other: 10
    # Sum: 148+72+61+43+14+10 = 348  ✓
    ethnicity_data = [
        # (Category, This Year Count, Last Year %)
        ('White',                        148,  0.4425),
        ('Hispanic/Latino',               72,  0.1987),
        ('Black/African American',        61,  0.1667),
        ('Asian',                         43,  0.1264),
        ('Two or More Races',             14,  0.0345),
        ('Other',                         10,  0.0259),
    ]

    # Write gender rows (rows 2-5, 4 gender groups)
    # Context says rows 2-8 for gender. We'll use rows 2-5 for the 4 groups,
    # leaving rows 6-8 if needed. But context says rows 2-8 for gender,
    # so let's use a section header at row 2 and data at rows 3-8 (6 rows).
    # Actually, context says rows 2-8 for gender groups (Male, Female, Non-Binary),
    # and rows 9-14 for ethnicity groups.
    # Let's use rows 2-8 for gender (add 2 more categories to fill rows 2-8).
    # Revised: Male, Female, Non-Binary, Prefer Not to Say = 4 rows.
    # Add "Gender Subtotal" and "Unknown" to reach row 8, or just use 4 rows + skip.
    # To match the context exactly: rows 2-8 is 7 rows for gender.
    # Let's split: Male, Female, Non-Binary, Prefer Not to Say, and add
    # Male (Full-Time), Male (Part-Time), Female (Full-Time) style.
    # Actually simpler: keep 4 groups but place total at row 6, leave 7-8 blank or add sub-cats.
    # The safest is to use exactly what makes the task meaningful:
    # rows 2-8 = gender section = 7 rows of gender data.

    # Revised gender data with 7 entries for rows 2-8:
    gender_data_7 = [
        ('Male',                            181,  0.5316),
        ('Male (Full-Time)',                162,  0.4770),
        ('Male (Part-Time)',                 19,  0.0546),
        ('Female',                          152,  0.4253),
        ('Female (Full-Time)',              138,  0.3908),
        ('Female (Part-Time)',               14,  0.0345),
        ('Non-Binary / Other',               15,  0.0431),
    ]
    # Check total: 162+19+138+14+15 = 348... no, these are just breakdowns.
    # Actually we keep B15 = 348. The sub-totals don't need to sum to 348.
    # The % formulas will use B2/$B$15 etc. so they're relative to total headcount.

    data_font = Font(name='Calibri', size=11)
    pct_format = '0.00%'
    int_format = '#,##0'
    align_center = Alignment(horizontal='center')
    align_right = Alignment(horizontal='right')

    # Write gender rows 2-8
    for row_offset, (cat, count, ly_pct) in enumerate(gender_data_7, 2):
        ws.cell(row=row_offset, column=1, value=cat).font = data_font
        count_cell = ws.cell(row=row_offset, column=2, value=count)
        count_cell.font = data_font
        count_cell.number_format = int_format
        count_cell.alignment = align_center
        # Column C (This Year %) — intentionally EMPTY (task must fill with formula)
        # Column D (Last Year %) — filled with prior year data
        ly_cell = ws.cell(row=row_offset, column=4, value=ly_pct)
        ly_cell.font = data_font
        ly_cell.number_format = pct_format
        ly_cell.alignment = align_center
        # Column E (Change) — intentionally EMPTY (task must fill with formula)

    # Write ethnicity rows 9-14
    for row_offset, (cat, count, ly_pct) in enumerate(ethnicity_data, 9):
        ws.cell(row=row_offset, column=1, value=cat).font = data_font
        count_cell = ws.cell(row=row_offset, column=2, value=count)
        count_cell.font = data_font
        count_cell.number_format = int_format
        count_cell.alignment = align_center
        # Column C — EMPTY (task fills)
        # Column D — prior year %
        ly_cell = ws.cell(row=row_offset, column=4, value=ly_pct)
        ly_cell.font = data_font
        ly_cell.number_format = pct_format
        ly_cell.alignment = align_center
        # Column E — EMPTY (task fills)

    # Row 15: Total row
    total_font = Font(name='Calibri', bold=True, size=11)
    thin = Side(style='thin', color='000000')
    top_border = Border(top=thin)

    ws.cell(row=15, column=1, value='Total').font = total_font
    total_b = ws.cell(row=15, column=2, value=348)
    total_b.font = total_font
    total_b.number_format = int_format
    total_b.alignment = align_center
    total_b.border = top_border

    # Add a section divider row between gender and ethnicity (visual)
    section_label_font = Font(name='Calibri', bold=True, size=10, color='FF595959')

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: Diversity Data')
    print('Rows 2-8: Gender breakdown (7 groups)')
    print('Rows 9-14: Ethnicity breakdown (6 groups)')
    print('Row 15: Total (B15=348)')
    print('Column C (This Year %) and E (Change): EMPTY — task must fill')
    print('Column D (Last Year %): filled with prior year percentages')
    print('No chart in initial state')


create_initial()
