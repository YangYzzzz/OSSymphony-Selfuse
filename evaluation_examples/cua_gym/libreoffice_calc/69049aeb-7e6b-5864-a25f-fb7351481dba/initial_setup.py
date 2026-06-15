"""
Initial Setup: Build a pivot table showing quarterly revenue trends
Task ID: calc_pivot_017
Domain: libreoffice_calc

Creates a Revenue sheet with 400 rows of invoice data spanning all 4 quarters of 2024.
Amounts are distributed to match target quarterly totals:
  Q1=82000, Q2=95000, Q3=88000, Q4=110000, Grand Total=375000
"""

import os
import shlex
import subprocess
import time
import random
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_017'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

# --- Realistic data pools ---
CLIENTS = [
    "Meridian Systems", "Atlas Logistics", "Pinnacle Consulting", "Vertex Analytics",
    "Horizon Media", "Cascade Technologies", "Silverline Partners", "NovaBridge Inc.",
    "Crestview Holdings", "Summit Digital", "Evergreen Solutions", "Bluewave Dynamics",
    "Ironclad Security", "TerraFirm Construction", "Luminary Health", "Redwood Financial",
    "ClearPath Energy", "Starforge Robotics", "Tidewater Shipping", "Keystone Advisors",
    "Quantum Leap Labs", "Greenfield Agriculture", "OceanView Resorts", "Granite Peak Mining",
    "FrostByte Computing", "SunRise Solar", "WildCard Entertainment", "MapleLeaf Logistics",
    "ThunderBolt Electric", "CoralReef Marine"
]

SERVICES = [
    "IT Consulting", "Cloud Migration", "Data Analytics", "Security Audit",
    "Software Development", "Network Setup", "Training Workshop", "System Integration",
    "Database Optimization", "UX Design", "Project Management", "API Development",
    "Compliance Review", "Infrastructure Setup", "Technical Support", "Mobile Development",
    "DevOps Services", "Business Intelligence", "Quality Assurance", "Digital Marketing"
]

def create_initial():
    random.seed(42)

    # Quarter targets
    q_targets = {1: 82000, 2: 95000, 3: 88000, 4: 110000}
    # 100 invoices per quarter
    rows_per_q = 100

    # Quarter date ranges
    q_ranges = {
        1: (date(2024, 1, 1), date(2024, 3, 31)),
        2: (date(2024, 4, 1), date(2024, 6, 30)),
        3: (date(2024, 7, 1), date(2024, 9, 30)),
        4: (date(2024, 10, 1), date(2024, 12, 31)),
    }

    # Generate amounts for each quarter that sum to target
    all_rows = []
    invoice_id = 1
    for q in range(1, 5):
        target = q_targets[q]
        start_d, end_d = q_ranges[q]
        days_range = (end_d - start_d).days

        # Generate 100 random amounts that sum to target
        raw = [random.uniform(50, 2000) for _ in range(rows_per_q)]
        raw_sum = sum(raw)
        amounts = [round(x / raw_sum * target, 2) for x in raw]
        # Fix rounding to hit exact target
        diff = round(target - sum(amounts), 2)
        amounts[-1] = round(amounts[-1] + diff, 2)

        for i in range(rows_per_q):
            inv_date = start_d + timedelta(days=random.randint(0, days_range))
            client = random.choice(CLIENTS)
            service = random.choice(SERVICES)
            all_rows.append((invoice_id, inv_date, client, service, amounts[i]))
            invoice_id += 1

    # Shuffle so they're not in quarter order (more realistic)
    random.shuffle(all_rows)
    # Re-assign sequential invoice IDs after shuffle
    all_rows = [(i + 1, r[1], r[2], r[3], r[4]) for i, r in enumerate(all_rows)]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revenue"

    # Headers
    headers = ["InvoiceID", "InvoiceDate", "Client", "Service", "Amount"]
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    white_font = Font(bold=True, size=11, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows
    for r, row_data in enumerate(all_rows, 2):
        ws.cell(row=r, column=1, value=row_data[0])  # InvoiceID
        ws.cell(row=r, column=2, value=row_data[1])  # InvoiceDate
        ws.cell(row=r, column=2).number_format = 'mm/dd/yyyy'
        ws.cell(row=r, column=3, value=row_data[2])  # Client
        ws.cell(row=r, column=4, value=row_data[3])  # Service
        ws.cell(row=r, column=5, value=row_data[4])  # Amount
        ws.cell(row=r, column=5).number_format = '#,##0.00'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 26
    ws.column_dimensions['D'].width = 24
    ws.column_dimensions['E'].width = 14

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Verify quarter totals
    q_sums = {1: 0, 2: 0, 3: 0, 4: 0}
    for row_data in all_rows:
        month = row_data[1].month
        if month <= 3:
            q_sums[1] += row_data[4]
        elif month <= 6:
            q_sums[2] += row_data[4]
        elif month <= 9:
            q_sums[3] += row_data[4]
        else:
            q_sums[4] += row_data[4]
    for q in range(1, 5):
        print(f'  Q{q}: {q_sums[q]:.2f} (target: {q_targets[q]})')
    print(f'  Total: {sum(q_sums.values()):.2f} (target: 375000)')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
