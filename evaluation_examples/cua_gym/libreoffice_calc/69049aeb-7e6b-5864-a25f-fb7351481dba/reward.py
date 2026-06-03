"""
Reward Script: Build a pivot table with quarterly revenue trends
Task ID: calc_pivot_017
Domain: libreoffice_calc
Scoring:
  C1 (0.20) - Pivot/summary sheet exists with quarterly data structure
  C2 (0.20) - Quarter labels (Q1-Q4) present in the pivot sheet
  C3 (0.40) - Quarterly revenue values match expected (Q1=82000, Q2=95000, Q3=88000, Q4=110000)
  C4 (0.20) - Grand total is 375000
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_017'

# Expected quarterly revenue values from task context
EXPECTED_QUARTERS = {
    'Q1': 82000,
    'Q2': 95000,
    'Q3': 88000,
    'Q4': 110000,
}
EXPECTED_GRAND_TOTAL = 375000
TOLERANCE = 500  # Allow small rounding differences


def find_pivot_sheet(wb):
    """Find a sheet that looks like a pivot/summary table (not 'Revenue')."""
    for sn in wb.sheetnames:
        if sn.lower() == 'revenue':
            continue
        ws = wb[sn]
        # Check if sheet has quarterly-looking data
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and 'q1' in cell.value.lower():
                    return ws
    # Fallback: return any non-Revenue sheet
    for sn in wb.sheetnames:
        if sn.lower() != 'revenue':
            return wb[sn]
    return None


def find_quarter_data(ws):
    """
    Scan the pivot sheet to find quarter labels and their associated values.
    Returns dict like {'Q1': value, 'Q2': value, ...} and grand_total.
    """
    quarter_values = {}
    grand_total = None

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for i, cell in enumerate(row):
            if cell.value and isinstance(cell.value, str):
                cell_text = cell.value.strip().upper()
                # Match Q1, Q2, Q3, Q4 in cell text
                q_match = re.search(r'\bQ([1-4])\b', cell_text)
                if q_match:
                    q_label = f'Q{q_match.group(1)}'
                    # Look for numeric value in the same row, to the right
                    for j in range(i + 1, len(row)):
                        val = row[j].value
                        if val is not None and isinstance(val, (int, float)):
                            quarter_values[q_label] = float(val)
                            break

                # Check for grand total
                if 'GRAND' in cell_text and 'TOTAL' in cell_text:
                    for j in range(i + 1, len(row)):
                        val = row[j].value
                        if val is not None and isinstance(val, (int, float)):
                            grand_total = float(val)
                            break
                elif cell_text == 'TOTAL':
                    for j in range(i + 1, len(row)):
                        val = row[j].value
                        if val is not None and isinstance(val, (int, float)):
                            # Only use as grand total if we haven't found one yet
                            if grand_total is None:
                                grand_total = float(val)
                            break

    return quarter_values, grand_total


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Also load with data_only to get computed values if formulas are used
    try:
        wb_data = openpyxl.load_workbook(file_path, data_only=True)
    except Exception:
        wb_data = None

    # Component 1: Pivot/summary sheet exists with quarterly structure (0.2 points)
    # This FAILS on initial (only Revenue sheet) and PASSES on golden (has Pivot sheet)
    try:
        pivot_ws = find_pivot_sheet(wb)
        pivot_ws_data = find_pivot_sheet(wb_data) if wb_data else None

        if pivot_ws is not None:
            # Verify it has some quarterly content (not just an empty sheet)
            quarterly_cell_count = sum(
                1 for row in pivot_ws.iter_rows(min_row=1, max_row=pivot_ws.max_row, max_col=pivot_ws.max_column)
                for cell in row
                if cell.value and isinstance(cell.value, str) and re.search(r'\bQ[1-4]\b', cell.value.upper())
            )

            if quarterly_cell_count > 0:
                print(f"PASS: Component 1 -- Pivot sheet '{pivot_ws.title}' exists with quarterly content (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 1 -- Sheet '{pivot_ws.title}' exists but has no quarterly labels (Q1-Q4)")
        else:
            print("FAIL: Component 1 -- No pivot/summary sheet found (only 'Revenue' exists)")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    if pivot_ws is None:
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Use data_only version if available for value checks
    check_ws = pivot_ws_data if pivot_ws_data else pivot_ws

    # Extract quarter data from the pivot sheet
    quarter_values, grand_total = find_quarter_data(check_ws)
    # Also try formula version if data_only didn't yield results
    if not quarter_values:
        quarter_values, grand_total = find_quarter_data(pivot_ws)

    print(f"  Found quarter values: {quarter_values}")
    print(f"  Found grand total: {grand_total}")

    # Component 2: Quarter labels present (0.2 points)
    # All four quarters Q1-Q4 should have labels in the pivot sheet
    try:
        found_quarters = set(quarter_values.keys())
        expected_set = {'Q1', 'Q2', 'Q3', 'Q4'}
        if found_quarters >= expected_set:
            print(f"PASS: Component 2 -- All 4 quarter labels found: {sorted(found_quarters)} (0.2 pts)")
            total_score += 0.2
        else:
            missing = expected_set - found_quarters
            print(f"FAIL: Component 2 -- Missing quarter labels: {missing}, found: {sorted(found_quarters)}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Quarterly revenue values correct (0.4 points, 0.1 per quarter)
    try:
        correct_quarters = 0
        for q_label, expected_val in EXPECTED_QUARTERS.items():
            actual_val = quarter_values.get(q_label)
            if actual_val is not None and abs(actual_val - expected_val) <= TOLERANCE:
                print(f"  PASS: {q_label} revenue = {actual_val} (expected {expected_val})")
                correct_quarters += 1
            else:
                print(f"  FAIL: {q_label} revenue = {actual_val} (expected {expected_val})")

        if correct_quarters > 0:
            c3_score = 0.1 * correct_quarters
            print(f"PASS: Component 3 -- {correct_quarters}/4 quarters correct ({c3_score} pts)")
            total_score += c3_score
        else:
            print("FAIL: Component 3 -- No quarterly values match expected")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Grand total correct (0.2 points)
    try:
        if grand_total is not None and abs(grand_total - EXPECTED_GRAND_TOTAL) <= TOLERANCE:
            print(f"PASS: Component 4 -- Grand total = {grand_total} (expected {EXPECTED_GRAND_TOTAL}) (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 4 -- Grand total = {grand_total} (expected {EXPECTED_GRAND_TOTAL})")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook for LibreOffice (save any unsaved edits)
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
