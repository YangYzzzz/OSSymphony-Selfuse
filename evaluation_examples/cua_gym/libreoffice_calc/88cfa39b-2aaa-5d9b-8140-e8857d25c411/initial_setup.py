"""
Initial Setup: Copy a sheet to another open LibreOffice Calc workbook
Task ID: calc_gsi_037
Domain: libreoffice_calc

Creates two workbooks:
  - master_report.xlsx  (source, contains "Executive Summary" sheet)
  - client_presentation.xlsx (target, does NOT contain "Executive Summary")
Opens both in LibreOffice Calc so the agent can use Move or Copy Sheet dialog.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_037'
MASTER_FILE = f'{WORKDIR}/master_report.xlsx'
CLIENT_FILE = f'{WORKDIR}/client_presentation.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_master_report():
    """Create master_report.xlsx with Executive Summary + other sheets."""
    wb = openpyxl.Workbook()

    # --- Sheet 1: Executive Summary ---
    ws_exec = wb.active
    ws_exec.title = 'Executive Summary'

    # Header styling
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    # Title row
    ws_exec.merge_cells('A1:E1')
    ws_exec['A1'] = 'Q1 2025 Executive Summary Report'
    ws_exec['A1'].font = Font(name='Arial', size=16, bold=True, color='2F5496')
    ws_exec['A1'].alignment = Alignment(horizontal='center')
    ws_exec.row_dimensions[1].height = 30

    # Headers in row 3
    exec_headers = ['Metric', 'Q4 2024', 'Q1 2025', 'Change', 'Status']
    for col, h in enumerate(exec_headers, 1):
        cell = ws_exec.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows
    exec_data = [
        ['Total Revenue', 2450000, 2785000, 13.7, 'On Track'],
        ['Operating Costs', 1120000, 1095000, -2.2, 'Improved'],
        ['Net Profit', 1330000, 1690000, 27.1, 'Exceeding'],
        ['Customer Count', 4520, 5130, 13.5, 'On Track'],
        ['Avg Revenue/Customer', 542, 543, 0.2, 'Stable'],
        ['Employee Headcount', 187, 203, 8.6, 'Growing'],
        ['Customer Satisfaction', 4.2, 4.5, 7.1, 'Improved'],
        ['Market Share %', 12.3, 14.1, 14.6, 'Exceeding'],
        ['New Product Lines', 3, 5, 66.7, 'Exceeding'],
        ['R&D Investment', 340000, 425000, 25.0, 'On Track'],
    ]
    for r, row_data in enumerate(exec_data, 4):
        for c, val in enumerate(row_data, 1):
            cell = ws_exec.cell(row=r, column=c, value=val)
            if c == 4:
                cell.number_format = '0.0%' if isinstance(val, float) else '0'
            elif c in (2, 3) and isinstance(val, (int, float)) and val > 1000:
                cell.number_format = '$#,##0'

    # Column widths
    ws_exec.column_dimensions['A'].width = 24
    ws_exec.column_dimensions['B'].width = 16
    ws_exec.column_dimensions['C'].width = 16
    ws_exec.column_dimensions['D'].width = 12
    ws_exec.column_dimensions['E'].width = 14

    # --- Sheet 2: Revenue Detail ---
    ws_rev = wb.create_sheet('Revenue Detail')
    rev_headers = ['Month', 'Product A', 'Product B', 'Product C', 'Services', 'Total']
    for col, h in enumerate(rev_headers, 1):
        cell = ws_rev.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    rev_data = [
        ['January', 185000, 142000, 98000, 210000, 635000],
        ['February', 192000, 156000, 105000, 225000, 678000],
        ['March', 210000, 168000, 118000, 248000, 744000],
        ['April', 198000, 149000, 112000, 235000, 694000],
        ['May', 215000, 171000, 125000, 252000, 763000],
        ['June', 228000, 183000, 131000, 268000, 810000],
        ['July', 205000, 160000, 119000, 241000, 725000],
        ['August', 219000, 175000, 128000, 258000, 780000],
        ['September', 232000, 188000, 135000, 270000, 825000],
        ['October', 241000, 195000, 140000, 278000, 854000],
        ['November', 248000, 201000, 145000, 285000, 879000],
        ['December', 255000, 210000, 152000, 295000, 912000],
    ]
    for r, row_data in enumerate(rev_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_rev.cell(row=r, column=c, value=val)
            if c >= 2:
                cell.number_format = '$#,##0'

    # --- Sheet 3: Regional Breakdown ---
    ws_reg = wb.create_sheet('Regional Breakdown')
    reg_headers = ['Region', 'Q1 Revenue', 'Q1 Costs', 'Q1 Profit', 'Headcount']
    for col, h in enumerate(reg_headers, 1):
        cell = ws_reg.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    reg_data = [
        ['North America', 1250000, 480000, 770000, 85],
        ['Europe', 680000, 295000, 385000, 52],
        ['Asia Pacific', 520000, 198000, 322000, 38],
        ['Latin America', 210000, 78000, 132000, 18],
        ['Middle East & Africa', 125000, 44000, 81000, 10],
    ]
    for r, row_data in enumerate(reg_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_reg.cell(row=r, column=c, value=val)
            if c in (2, 3, 4):
                cell.number_format = '$#,##0'

    wb.save(MASTER_FILE)
    print(f'Created: {MASTER_FILE}')


def create_client_presentation():
    """Create client_presentation.xlsx WITHOUT Executive Summary sheet."""
    wb = openpyxl.Workbook()

    # --- Sheet 1: Project Overview ---
    ws_proj = wb.active
    ws_proj.title = 'Project Overview'

    ws_proj.merge_cells('A1:D1')
    ws_proj['A1'] = 'Client Partnership - Horizon Technologies'
    ws_proj['A1'].font = Font(name='Arial', size=14, bold=True, color='1F4E79')
    ws_proj['A1'].alignment = Alignment(horizontal='center')

    proj_headers = ['Deliverable', 'Status', 'Due Date', 'Owner']
    for col, h in enumerate(proj_headers, 1):
        cell = ws_proj.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFD9E2F3', end_color='FFD9E2F3', fill_type='solid')

    proj_data = [
        ['Phase 1 - Discovery', 'Complete', '2025-01-15', 'Sarah Chen'],
        ['Phase 2 - Design', 'Complete', '2025-02-28', 'Marcus Johnson'],
        ['Phase 3 - Development', 'In Progress', '2025-04-30', 'Anika Patel'],
        ['Phase 4 - Testing', 'Pending', '2025-06-15', 'James O\'Brien'],
        ['Phase 5 - Deployment', 'Pending', '2025-07-31', 'Li Wei'],
        ['Phase 6 - Support', 'Not Started', '2025-09-30', 'Rachel Kim'],
    ]
    for r, row_data in enumerate(proj_data, 4):
        for c, val in enumerate(row_data, 1):
            ws_proj.cell(row=r, column=c, value=val)

    ws_proj.column_dimensions['A'].width = 28
    ws_proj.column_dimensions['B'].width = 16
    ws_proj.column_dimensions['C'].width = 14
    ws_proj.column_dimensions['D'].width = 18

    # --- Sheet 2: Budget ---
    ws_budget = wb.create_sheet('Budget')
    budget_headers = ['Category', 'Allocated', 'Spent', 'Remaining']
    for col, h in enumerate(budget_headers, 1):
        cell = ws_budget.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    budget_data = [
        ['Personnel', 450000, 185000, 265000],
        ['Infrastructure', 120000, 72000, 48000],
        ['Software Licenses', 85000, 85000, 0],
        ['Travel & Meetings', 35000, 12000, 23000],
        ['Training', 25000, 8000, 17000],
        ['Contingency', 60000, 0, 60000],
        ['Marketing', 40000, 18500, 21500],
        ['External Consultants', 95000, 42000, 53000],
    ]
    for r, row_data in enumerate(budget_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_budget.cell(row=r, column=c, value=val)
            if c >= 2:
                cell.number_format = '$#,##0'

    ws_budget.column_dimensions['A'].width = 22
    ws_budget.column_dimensions['B'].width = 14
    ws_budget.column_dimensions['C'].width = 14
    ws_budget.column_dimensions['D'].width = 14

    # --- Sheet 3: Timeline ---
    ws_timeline = wb.create_sheet('Timeline')
    timeline_headers = ['Milestone', 'Start Date', 'End Date', 'Duration (days)', 'Dependencies']
    for col, h in enumerate(timeline_headers, 1):
        cell = ws_timeline.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    timeline_data = [
        ['Kickoff Meeting', '2025-01-06', '2025-01-06', 1, 'None'],
        ['Requirements Gathering', '2025-01-07', '2025-01-31', 25, 'Kickoff'],
        ['Architecture Review', '2025-02-03', '2025-02-14', 12, 'Requirements'],
        ['UI/UX Design', '2025-02-10', '2025-03-07', 26, 'Architecture'],
        ['Backend Development', '2025-03-03', '2025-04-18', 47, 'Architecture'],
        ['Frontend Development', '2025-03-10', '2025-04-25', 47, 'UI/UX Design'],
        ['Integration Testing', '2025-04-28', '2025-05-23', 26, 'Backend, Frontend'],
        ['UAT', '2025-05-26', '2025-06-13', 19, 'Integration Testing'],
        ['Go-Live Prep', '2025-06-16', '2025-07-11', 26, 'UAT'],
        ['Production Deployment', '2025-07-14', '2025-07-25', 12, 'Go-Live Prep'],
    ]
    for r, row_data in enumerate(timeline_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_timeline.cell(row=r, column=c, value=val)

    ws_timeline.column_dimensions['A'].width = 24
    ws_timeline.column_dimensions['B'].width = 14
    ws_timeline.column_dimensions['C'].width = 14
    ws_timeline.column_dimensions['D'].width = 16
    ws_timeline.column_dimensions['E'].width = 22

    wb.save(CLIENT_FILE)
    print(f'Created: {CLIENT_FILE}')


def main():
    create_master_report()
    create_client_presentation()

    # GUI-ready: open both workbooks in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{MASTER_FILE}"', delay_sec=3.0)
    launch_gui(f'libreoffice --calc "{CLIENT_FILE}"', delay_sec=2.0)
    print('GUI_READY: launched both workbooks with DISPLAY=:0')


main()
