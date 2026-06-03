"""
Reward Script: Copy 'Executive Summary' sheet from master_report.xlsx to client_presentation.xlsx
Task ID: calc_gsi_037
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): 'Executive Summary' sheet exists in client_presentation.xlsx
  Component 2 (0.35): Key header and data values match the source sheet
  Component 3 (0.25): All 13 rows of data are present (complete copy)
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_037'

# Expected key values from the Executive Summary sheet in master_report.xlsx
# These are the ground truth values that should appear in the copied sheet.
EXPECTED_HEADER = 'Q1 2025 Executive Summary Report'
EXPECTED_COL_HEADERS = ['Metric', 'Q4 2024', 'Q1 2025', 'Change', 'Status']
EXPECTED_KEY_DATA = {
    'A4': 'Total Revenue',
    'B4': 2450000,
    'C4': 2785000,
    'A6': 'Net Profit',
    'C6': 1690000,
    'A13': 'R&D Investment',
    'C13': 425000,
    'E4': 'On Track',
    'E6': 'Exceeding',
}


def verify_task():
    """
    Verify that the 'Executive Summary' sheet was copied from master_report.xlsx
    into client_presentation.xlsx with all data intact.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    client_path = f'{WORKDIR}/client_presentation.xlsx'

    try:
        wb = openpyxl.load_workbook(client_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load {client_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: 'Executive Summary' sheet exists in client_presentation.xlsx (0.4 points)
    # This is the primary task-introduced change: initial has 3 sheets, golden has 4.
    try:
        if 'Executive Summary' in wb.sheetnames:
            print(f"PASS: Component 1 — 'Executive Summary' sheet found in client_presentation.xlsx (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — 'Executive Summary' sheet not found. Sheets: {wb.sheetnames}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Key header and data values match the source sheet (0.35 points)
    # Verifies that the copied sheet has the correct content, not just the name.
    try:
        if 'Executive Summary' not in wb.sheetnames:
            print("FAIL: Component 2 — Cannot check data, sheet does not exist")
        else:
            ws = wb['Executive Summary']
            matches = 0
            total_checks = 0

            # Check title cell A1
            total_checks += 1
            if ws['A1'].value == EXPECTED_HEADER:
                matches += 1
            else:
                print(f"  DATA: A1 expected '{EXPECTED_HEADER}', found '{ws['A1'].value}'")

            # Check column headers in row 3
            for col_idx, expected_hdr in enumerate(EXPECTED_COL_HEADERS, 1):
                total_checks += 1
                actual = ws.cell(row=3, column=col_idx).value
                if actual == expected_hdr:
                    matches += 1
                else:
                    print(f"  DATA: Row 3 Col {col_idx} expected '{expected_hdr}', found '{actual}'")

            # Check key data cells
            for coord, expected_val in EXPECTED_KEY_DATA.items():
                total_checks += 1
                actual = ws[coord].value
                if isinstance(expected_val, (int, float)):
                    try:
                        if actual is not None and abs(float(actual) - expected_val) < 0.01:
                            matches += 1
                        else:
                            print(f"  DATA: {coord} expected {expected_val}, found {actual}")
                    except (ValueError, TypeError):
                        print(f"  DATA: {coord} expected {expected_val}, found {actual} (type mismatch)")
                else:
                    if str(actual).strip() == str(expected_val).strip():
                        matches += 1
                    else:
                        print(f"  DATA: {coord} expected '{expected_val}', found '{actual}'")

            match_ratio = matches / total_checks if total_checks > 0 else 0
            if match_ratio >= 0.8:
                print(f"PASS: Component 2 — {matches}/{total_checks} key values match ({match_ratio:.0%}) (0.35 pts)")
                total_score += 0.35
            elif match_ratio >= 0.5:
                partial = round(0.35 * match_ratio, 2)
                print(f"PARTIAL: Component 2 — {matches}/{total_checks} key values match ({match_ratio:.0%}) ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 2 — Only {matches}/{total_checks} key values match ({match_ratio:.0%})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: All rows present — complete copy with 13 rows of data (0.25 points)
    # The source sheet has data in rows 1-13 (A1:E13). A complete copy should have all of them.
    try:
        if 'Executive Summary' not in wb.sheetnames:
            print("FAIL: Component 3 — Cannot check row count, sheet does not exist")
        else:
            ws = wb['Executive Summary']
            # Count non-empty rows (at least one cell with a value)
            non_empty_rows = 0
            for row in ws.iter_rows(min_row=1, max_row=20, max_col=5):
                if any(cell.value is not None for cell in row):
                    non_empty_rows += 1

            if non_empty_rows >= 12:
                # 12 out of 13 rows have data (row 2 is blank separator)
                print(f"PASS: Component 3 — {non_empty_rows} non-empty rows found (expected 12+) (0.25 pts)")
                total_score += 0.25
            elif non_empty_rows >= 6:
                partial = round(0.25 * (non_empty_rows / 12), 2)
                print(f"PARTIAL: Component 3 — {non_empty_rows} non-empty rows (expected 12+) ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 3 — Only {non_empty_rows} non-empty rows found (expected 12+)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint
import os
client_path = f'{WORKDIR}/client_presentation.xlsx'
if not os.path.exists(client_path):
    print(f"File not found: {client_path}")
    print("REWARD: 0.0")
else:
    verify_task()
