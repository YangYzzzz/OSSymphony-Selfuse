"""
Initial Setup: Configure print scaling for a wide ledger sheet
Task ID: calc_mcp_081
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_081'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ledger'

    # --- Headers (A-R = 18 columns) ---
    headers = [
        'Transaction ID',    # A
        'Date',              # B
        'Account Code',      # C
        'Account Name',      # D
        'Department',        # E
        'Category',          # F
        'Description',       # G
        'Debit ($)',         # H
        'Credit ($)',        # I
        'Balance ($)',       # J
        'Currency',          # K
        'Exchange Rate',     # L
        'Local Amount ($)',  # M
        'Vendor / Payee',    # N
        'Invoice Number',    # O
        'Payment Method',    # P
        'Approved By',       # Q
        'Notes',             # R
    ]
    header_font = Font(bold=True, size=11, name='Calibri')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_font_white = Font(bold=True, size=11, name='Calibri', color='FFFFFF')
    thin = Side(style='thin', color='000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = header_border

    # --- Data rows 2-50 (49 rows) ---
    departments = ['Finance', 'Engineering', 'Marketing', 'Operations', 'HR', 'Sales', 'Legal', 'IT']
    categories = ['Salary', 'Supplies', 'Travel', 'Software', 'Consulting', 'Utilities', 'Insurance', 'Rent', 'Equipment', 'Training']
    accounts = [
        ('1001', 'Cash'), ('1010', 'Accounts Receivable'), ('2001', 'Accounts Payable'),
        ('4001', 'Sales Revenue'), ('5001', 'Cost of Goods'), ('6001', 'Office Supplies'),
        ('6010', 'Travel Expenses'), ('6020', 'Software Licenses'), ('6030', 'Professional Fees'),
        ('7001', 'Depreciation'),
    ]
    vendors = [
        'Acme Corp', 'TechVision Inc', 'Global Logistics LLC', 'Pinnacle Consulting',
        'CloudNine Software', 'Metro Utilities', 'SafeGuard Insurance', 'Apex Equipment',
        'BrightPath Training', 'Summit Supplies Co', 'NetWave Solutions', 'BlueRidge Partners',
    ]
    approvers = ['Sarah Chen', 'Marcus Johnson', 'Priya Patel', 'David Kim', 'Elena Rodriguez', 'James Wright']
    payment_methods = ['Wire Transfer', 'ACH', 'Check', 'Corporate Card', 'EFT']
    currencies = ['USD', 'USD', 'USD', 'EUR', 'GBP', 'USD', 'CAD', 'USD', 'USD', 'JPY']
    notes_options = [
        'Monthly recurring', 'One-time purchase', 'Quarterly payment', 'Annual renewal',
        'Project-based', 'Approved by VP', 'Urgent request', 'Standard procurement',
        'Budget allocation Q2', 'Pending review', 'Expedited processing', '',
    ]

    import random
    random.seed(42)

    running_balance = 125000.00
    for r in range(2, 51):
        idx = r - 2
        txn_id = f'TXN-2025-{idx + 1001:04d}'
        day = (idx * 3) % 28 + 1
        month = (idx // 10) % 12 + 1
        date_str = f'2025-{month:02d}-{day:02d}'
        acct = accounts[idx % len(accounts)]
        dept = departments[idx % len(departments)]
        cat = categories[idx % len(categories)]
        desc = f'{cat} - {dept} department allocation'
        debit = round(random.uniform(500, 25000), 2) if idx % 3 != 2 else 0
        credit = round(random.uniform(500, 25000), 2) if idx % 3 == 2 else 0
        running_balance += credit - debit
        curr = currencies[idx % len(currencies)]
        rate = 1.0 if curr == 'USD' else round(random.uniform(0.7, 1.5), 4)
        local_amt = round((debit if debit else credit) * rate, 2)
        vendor = vendors[idx % len(vendors)]
        inv = f'INV-{random.randint(10000, 99999)}'
        pm = payment_methods[idx % len(payment_methods)]
        approver = approvers[idx % len(approvers)]
        note = notes_options[idx % len(notes_options)]

        row_data = [
            txn_id, date_str, acct[0], acct[1], dept, cat, desc,
            debit, credit, round(running_balance, 2), curr, rate,
            local_amt, vendor, inv, pm, approver, note
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = header_border
            # Number formatting
            if c in (8, 9, 10, 13):
                cell.number_format = '$#,##0.00'
            elif c == 12:
                cell.number_format = '0.0000'

    # Set column widths to make sheet realistically wide
    col_widths = {
        'A': 16, 'B': 12, 'C': 14, 'D': 22, 'E': 14, 'F': 14, 'G': 35,
        'H': 14, 'I': 14, 'J': 14, 'K': 10, 'L': 14, 'M': 16,
        'N': 22, 'O': 16, 'P': 18, 'Q': 18, 'R': 24,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 30

    # DO NOT set any print scaling - leave default
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
