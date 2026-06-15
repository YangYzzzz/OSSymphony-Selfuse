"""
Initial Setup: Project status tracker with header row and 40 project entries
Task ID: calc_gsd_013
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_013'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projects"

    # --- Row 1: Merged title ---
    ws.merge_cells("A1:H1")
    ws["A1"] = "Project Status Dashboard 2024"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # --- Row 2: Column headers ---
    headers = [
        "Project ID", "Project Name", "Owner", "Start Date",
        "End Date", "% Complete", "Status", "Priority"
    ]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF1F4E79", end_color="FF1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[2].height = 24

    # --- Column widths ---
    col_widths = {"A": 12, "B": 32, "C": 20, "D": 14, "E": 14, "F": 14, "G": 16, "H": 12}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # --- Rows 3-42: 40 project entries ---
    projects = [
        ["PRJ-001", "Cloud Migration Phase 1", "Sarah Chen", "2024-01-15", "2024-04-30", 100, "Complete", "High"],
        ["PRJ-002", "Mobile App Redesign", "Marcus Johnson", "2024-02-01", "2024-06-15", 72, "In Progress", "High"],
        ["PRJ-003", "Data Warehouse Upgrade", "Elena Rodriguez", "2024-01-20", "2024-05-31", 100, "Complete", "Medium"],
        ["PRJ-004", "Customer Portal v2.0", "David Kim", "2024-03-01", "2024-07-15", 15, "Blocked", "High"],
        ["PRJ-005", "ERP Integration", "Priya Patel", "2024-02-10", "2024-08-30", 45, "In Progress", "Critical"],
        ["PRJ-006", "Security Audit 2024", "James O'Brien", "2024-01-05", "2024-03-15", 100, "Complete", "Critical"],
        ["PRJ-007", "AI Chatbot Development", "Lisa Wang", "2024-03-15", "2024-09-30", 30, "In Progress", "Medium"],
        ["PRJ-008", "Office Relocation IT Setup", "Robert Taylor", "2024-04-01", "2024-06-30", 0, "Not Started", "Low"],
        ["PRJ-009", "Payment Gateway Migration", "Aisha Mohammed", "2024-02-15", "2024-05-15", 88, "In Progress", "High"],
        ["PRJ-010", "Legacy System Decommission", "Carlos Mendoza", "2024-01-10", "2024-12-31", 22, "Blocked", "Medium"],
        ["PRJ-011", "Employee Onboarding Portal", "Jennifer Adams", "2024-03-20", "2024-07-30", 100, "Complete", "Medium"],
        ["PRJ-012", "Network Infrastructure Refresh", "Michael Brown", "2024-04-15", "2024-10-15", 10, "In Progress", "High"],
        ["PRJ-013", "Compliance Dashboard", "Fatima Al-Hassan", "2024-02-28", "2024-06-30", 65, "In Progress", "Critical"],
        ["PRJ-014", "Vendor Management System", "Thomas Fischer", "2024-05-01", "2024-11-30", 0, "Not Started", "Low"],
        ["PRJ-015", "Disaster Recovery Plan Update", "Samantha Lee", "2024-01-25", "2024-04-15", 100, "Complete", "Critical"],
        ["PRJ-016", "CRM Data Cleanup", "Daniel Nguyen", "2024-03-10", "2024-05-31", 50, "Blocked", "Medium"],
        ["PRJ-017", "API Gateway Implementation", "Rachel Green", "2024-04-20", "2024-08-15", 35, "In Progress", "High"],
        ["PRJ-018", "Warehouse Automation Phase 2", "Omar Khalil", "2024-02-05", "2024-09-15", 60, "In Progress", "High"],
        ["PRJ-019", "Brand Website Refresh", "Emily Parker", "2024-05-15", "2024-08-30", 5, "In Progress", "Medium"],
        ["PRJ-020", "SOC 2 Certification", "Kevin Matthews", "2024-01-08", "2024-06-30", 100, "Complete", "Critical"],
        ["PRJ-021", "Inventory Tracking System", "Nadia Volkov", "2024-03-25", "2024-07-31", 40, "In Progress", "Medium"],
        ["PRJ-022", "Cloud Cost Optimization", "Brian Walsh", "2024-04-10", "2024-06-15", 78, "In Progress", "High"],
        ["PRJ-023", "Customer Feedback Analytics", "Sophie Martin", "2024-05-01", "2024-10-30", 0, "Not Started", "Low"],
        ["PRJ-024", "Single Sign-On Rollout", "Alex Petrov", "2024-02-20", "2024-05-20", 100, "Complete", "High"],
        ["PRJ-025", "Supply Chain Dashboard", "Maria Gonzalez", "2024-03-05", "2024-08-15", 55, "In Progress", "Medium"],
        ["PRJ-026", "IT Help Desk Automation", "Jason Park", "2024-04-25", "2024-09-30", 20, "Blocked", "Medium"],
        ["PRJ-027", "Data Lake Migration", "Anita Sharma", "2024-01-30", "2024-07-15", 82, "In Progress", "High"],
        ["PRJ-028", "Mobile POS System", "Christopher Davis", "2024-05-10", "2024-11-15", 0, "Not Started", "Medium"],
        ["PRJ-029", "Regulatory Reporting Tool", "Laura Bennett", "2024-02-12", "2024-06-30", 100, "Complete", "Critical"],
        ["PRJ-030", "DevOps Pipeline Overhaul", "Ryan Cooper", "2024-03-18", "2024-08-30", 48, "In Progress", "High"],
        ["PRJ-031", "Customer Loyalty Platform", "Hannah Wilson", "2024-04-05", "2024-10-31", 25, "In Progress", "Medium"],
        ["PRJ-032", "Email System Migration", "Victor Santos", "2024-01-15", "2024-04-30", 100, "Complete", "High"],
        ["PRJ-033", "Quality Assurance Framework", "Diana Reeves", "2024-05-20", "2024-12-15", 8, "In Progress", "Low"],
        ["PRJ-034", "Financial Reconciliation Bot", "Gregory Hall", "2024-03-12", "2024-07-15", 62, "Blocked", "High"],
        ["PRJ-035", "Multi-Region Deployment", "Yuki Tanaka", "2024-02-25", "2024-09-30", 38, "In Progress", "Critical"],
        ["PRJ-036", "Asset Management System", "Patricia Murphy", "2024-04-15", "2024-10-15", 0, "Not Started", "Medium"],
        ["PRJ-037", "Cybersecurity Training Portal", "Ahmed Hassan", "2024-01-20", "2024-05-15", 100, "Complete", "High"],
        ["PRJ-038", "Business Intelligence Suite", "Olivia Stewart", "2024-03-28", "2024-11-30", 32, "In Progress", "Medium"],
        ["PRJ-039", "Automated Testing Framework", "Nathan Brooks", "2024-05-05", "2024-09-15", 18, "Blocked", "High"],
        ["PRJ-040", "Customer Data Platform", "Isabella Cruz", "2024-02-08", "2024-08-31", 70, "In Progress", "Critical"],
    ]

    data_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for r, row_data in enumerate(projects, 3):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = data_border
            cell.font = Font(name="Calibri", size=11)
            if c == 4 or c == 5:  # Date columns
                cell.number_format = 'yyyy-mm-dd'
            elif c == 6:  # % Complete
                cell.alignment = Alignment(horizontal="center")
                cell.number_format = '0"%"'

    # NO freeze panes (task requires agent to add them)
    # NO conditional formatting (task requires agent to add it)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
