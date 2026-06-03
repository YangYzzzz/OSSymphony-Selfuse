"""
Reward Script: Protect 'Formulas' sheet without password, lock all cells, hide formula cells C2:C50
Task ID: calc_ps_037
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): Sheet protection is enabled
  Component 2 (0.15): Protection has no password
  Component 3 (0.35): All formula cells C2:C50 are hidden (proportional)
  Component 4 (0.15): All cells remain locked with protection active
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_037'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Formulas' sheet must exist
    if 'Formulas' not in wb.sheetnames:
        print(f"CRITICAL: 'Formulas' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Formulas']

    # Component 1: Sheet protection is enabled (0.35 points)
    # Initial: protection.sheet == False; Golden: protection.sheet == True
    try:
        if ws.protection.sheet:
            print(f"PASS: Component 1 -- Sheet is protected (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 1 -- Sheet is NOT protected (protection.sheet={ws.protection.sheet})")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Protection has no password (0.15 points)
    # Only meaningful when sheet IS protected. Initial has no protection -> fails gate.
    try:
        if ws.protection.sheet:
            pw = ws.protection.password
            if pw is None or pw == '':
                print(f"PASS: Component 2 -- No password on protection (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 2 -- Password is set: '{pw}'")
        else:
            print(f"FAIL: Component 2 -- Sheet not protected, cannot verify password absence")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: All formula cells C2:C50 are hidden (0.35 points, proportional)
    # Initial: all hidden=False; Golden: all hidden=True
    try:
        hidden_count = 0
        total_cells = 49  # C2 through C50
        for r in range(2, 51):
            cell = ws.cell(row=r, column=3)
            if cell.protection.hidden:
                hidden_count += 1

        if hidden_count == total_cells:
            print(f"PASS: Component 3 -- All {total_cells} cells in C2:C50 are hidden (0.35 pts)")
            total_score += 0.35
        elif hidden_count > 0:
            partial = round(0.35 * (hidden_count / total_cells), 4)
            print(f"PARTIAL: Component 3 -- {hidden_count}/{total_cells} cells hidden ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 -- No cells in C2:C50 are hidden (0/49)")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: All cells remain locked with protection active (0.15 points)
    # Cells are locked in initial too, but without protection it's meaningless.
    # Gate on protection being active so initial_env scores 0.
    try:
        if ws.protection.sheet:
            # Spot-check a representative set of cells across all columns and rows
            check_coords = ['A1', 'B1', 'C1', 'A2', 'B2', 'C2',
                            'A10', 'B10', 'C10', 'A25', 'B25', 'C25',
                            'A50', 'B50', 'C50']
            unlocked_cells = []
            for coord in check_coords:
                cell = ws[coord]
                if not cell.protection.locked:
                    unlocked_cells.append(coord)

            if len(unlocked_cells) == 0:
                print(f"PASS: Component 4 -- All checked cells are locked with protection active (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 4 -- Unlocked cells found: {unlocked_cells}")
        else:
            print(f"FAIL: Component 4 -- Sheet not protected, locked cells have no effect")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
