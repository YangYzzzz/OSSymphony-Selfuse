"""
Initial Setup: Protect the 'Formulas' sheet without a password.
Task ID: calc_ps_037
Domain: libreoffice_calc

Creates a workbook with a 'Formulas' sheet containing X, Y, and Result columns.
C2:C50 contain formulas. All cells locked, none hidden. Sheet unprotected.
"""

import os
import shlex
import subprocess
import time
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Protection

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_037'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Formulas'

    # --- Headers ---
    headers = ['X', 'Y', 'Result']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # --- Data rows (A2:B50 numeric, C2:C50 formulas) ---
    random.seed(42)  # reproducible
    offsets = [
        3.5, -2.1, 7.8, 0.5, -1.3, 4.2, 6.0, -3.7, 2.9, 1.1,
        -0.8, 5.4, 3.3, -4.6, 8.2, 0.7, -2.5, 6.1, 1.9, -3.0,
        4.8, 2.2, -1.7, 7.3, 0.3, -5.1, 3.6, 9.0, -0.4, 2.7,
        -6.2, 1.5, 4.4, -2.8, 5.9, 0.1, -3.4, 7.6, 2.0, -1.0,
        8.5, 3.1, -4.3, 6.7, 0.9, -2.2, 5.0, 1.4, -0.6
    ]

    for i in range(49):  # rows 2 through 50
        row = i + 2
        x_val = round(random.uniform(1, 100), 1)
        y_val = round(random.uniform(1, 50), 1)
        ws.cell(row=row, column=1, value=x_val)
        ws.cell(row=row, column=2, value=y_val)
        offset = offsets[i]
        sign = '+' if offset >= 0 else ''
        ws.cell(row=row, column=3, value=f'=A{row}*B{row}{sign}{offset}')

    # --- All cells locked, none hidden (openpyxl default is locked=True, hidden=False) ---
    # Ensure explicitly: set protection on all used cells
    for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=3):
        for cell in row:
            cell.protection = Protection(locked=True, hidden=False)

    # --- Sheet is NOT protected ---
    ws.protection.sheet = False

    # --- Column widths ---
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
