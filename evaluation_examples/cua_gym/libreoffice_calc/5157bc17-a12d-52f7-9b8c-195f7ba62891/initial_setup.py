"""
Initial Setup: Calculate straight-line depreciation for fixed assets
Task ID: calc_fin_asset_depreciation_031
Domain: libreoffice_calc

Creates a workbook with:
  - Sheet 'Assets': 11 assets with Cost, Salvage Value, Useful Life filled in;
    Annual Depr column (E) is EMPTY (to be filled by agent)
  - Sheet 'DepreciationSchedule': completely empty (to be filled by agent)
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_asset_depreciation_031'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # -------------------------------------------------------------------------
    # Sheet 1: Assets
    # -------------------------------------------------------------------------
    ws = wb.active
    ws.title = 'Assets'

    # Headers row 1: A=Asset, B=Cost, C=Salvage Value, D=Useful Life (Years), E=Annual Depr
    headers = ['Asset', 'Cost', 'Salvage Value', 'Useful Life (Years)', 'Annual Depr']
    header_font = Font(bold=True, name='Calibri', size=11)
    header_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 11 realistic fixed assets — E column intentionally left empty
    assets = [
        # (Asset Name,           Cost,       Salvage,   Useful Life)
        ('CNC Milling Machine',  185000,     15000,     10),
        ('Delivery Truck #1',     62500,      5000,      7),
        ('Office Building Wing',  540000,     80000,     25),
        ('Industrial HVAC Unit',   38400,      3200,      8),
        ('Server Rack Array',       72000,      6000,     5),
        ('Forklift — Model FX3',   47300,      4000,     12),
        ('Laser Cutting System',  230000,     20000,     15),
        ('Company Sedan Fleet',    94000,      8000,      6),
        ('Solar Panel Array',     115000,     10000,     20),
        ('Packaging Line Robot',  162000,     12000,     10),
        ('Digital Printing Press', 98500,      9500,      8),
    ]

    for r, (name, cost, salvage, life) in enumerate(assets, 2):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=cost).number_format = '$#,##0.00'
        ws.cell(row=r, column=3, value=salvage).number_format = '$#,##0.00'
        ws.cell(row=r, column=4, value=life)
        # Column E (Annual Depr) intentionally left empty

    # Apply currency format retroactively for col B and C
    for r in range(2, 13):
        ws.cell(row=r, column=2).number_format = '$#,##0.00'
        ws.cell(row=r, column=3).number_format = '$#,##0.00'

    # Column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 16

    ws.row_dimensions[1].height = 22

    # Freeze pane below header
    ws.freeze_panes = 'A2'

    # -------------------------------------------------------------------------
    # Sheet 2: DepreciationSchedule — completely empty
    # -------------------------------------------------------------------------
    ws2 = wb.create_sheet('DepreciationSchedule')
    # Leave completely empty as per task context

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('  Sheet "Assets": headers + 11 assets, E column empty, no totals row')
    print('  Sheet "DepreciationSchedule": empty')


create_initial()
