"""
Reward Script: Calculate straight-line depreciation for fixed assets
Task ID: calc_fin_asset_depreciation_031
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.25): Assets sheet E2:E12 have depreciation formulas =(Bx-Cx)/Dx
  - Component 2 (0.25): Assets sheet row 13 has SUM totals row, bold formatted
  - Component 3 (0.25): DepreciationSchedule has headers and year-by-year data with correct formulas
  - Component 4 (0.25): DepreciationSchedule has a stacked bar chart with 2 series
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fin_asset_depreciation_031'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # ---- Component 1: Assets E2:E12 annual depreciation formulas (0.25 points) ----
    # Task asks to fill E column with =(Bx-Cx)/Dx formula for each of 11 assets
    # Initial file has E2:E12 empty; golden file has formulas there
    try:
        if 'Assets' not in wb.sheetnames:
            print("FAIL: Component 1 — 'Assets' sheet not found")
        else:
            ws_assets = wb['Assets']
            formulas_found = 0
            formulas_correct = 0
            for row in range(2, 13):  # rows 2 through 12 (11 assets)
                val = ws_assets.cell(row=row, column=5).value  # Column E
                if val is not None:
                    formulas_found += 1
                    # Check formula pattern: =(Brow-Crow)/Drow
                    if isinstance(val, str):
                        # Normalize: remove spaces, uppercase
                        normalized = val.upper().replace(' ', '')
                        expected = f'=(B{row}-C{row})/D{row}'
                        if normalized == expected:
                            formulas_correct += 1

            if formulas_correct == 11:
                print(f"PASS: Component 1 — All 11 depreciation formulas in E2:E12 are correct ({formulas_correct}/11)")
                total_score += 0.25
            elif formulas_found >= 1 and formulas_correct >= 6:
                print(f"PARTIAL: Component 1 — {formulas_correct}/11 depreciation formulas correct in E column")
                total_score += 0.10
            else:
                print(f"FAIL: Component 1 — Expected 11 formulas in E2:E12, found {formulas_found} values, {formulas_correct} correct")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ---- Component 2: Assets row 13 totals row with bold formatting (0.25 points) ----
    # Task asks for Row 13: totals for B, C, E columns, bold
    # Initial file has no row 13; golden adds it
    try:
        if 'Assets' not in wb.sheetnames:
            print("FAIL: Component 2 — 'Assets' sheet not found")
        else:
            ws_assets = wb['Assets']
            a13 = ws_assets.cell(row=13, column=1)
            b13 = ws_assets.cell(row=13, column=2)
            c13 = ws_assets.cell(row=13, column=3)
            e13 = ws_assets.cell(row=13, column=5)

            # Check that row 13 has total label and SUM formulas
            has_label = a13.value is not None and str(a13.value).strip() != ''
            has_b_sum = b13.value is not None and isinstance(b13.value, str) and 'SUM' in b13.value.upper()
            has_c_sum = c13.value is not None and isinstance(c13.value, str) and 'SUM' in c13.value.upper()
            has_e_sum = e13.value is not None and isinstance(e13.value, str) and 'SUM' in e13.value.upper()

            # Check bold for at least B13 or A13
            is_bold = (a13.font.bold is True) or (b13.font.bold is True)

            if has_label and has_b_sum and has_c_sum and has_e_sum and is_bold:
                print(f"PASS: Component 2 — Row 13 has total label='{a13.value}', B13={b13.value}, C13={c13.value}, E13={e13.value}, bold={is_bold}")
                total_score += 0.25
            elif has_b_sum and has_c_sum and has_e_sum:
                print(f"PARTIAL: Component 2 — Row 13 has SUM formulas but label or bold missing (label={has_label}, bold={is_bold})")
                total_score += 0.12
            else:
                print(f"FAIL: Component 2 — Row 13 missing totals: label={has_label}, B_sum={has_b_sum}, C_sum={has_c_sum}, E_sum={has_e_sum}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---- Component 3: DepreciationSchedule headers + data rows with correct formulas (0.25 points) ----
    # Task asks for Year/BegBV/Depr/AccDepr/EndingBV headers, and yearly data rows
    # Initial sheet is empty; golden fills it with 25 years of data
    try:
        if 'DepreciationSchedule' not in wb.sheetnames:
            print("FAIL: Component 3 — 'DepreciationSchedule' sheet not found")
        else:
            ws_depr = wb['DepreciationSchedule']

            # Check headers in row 1
            expected_headers = ['Year', 'Beginning BV', 'Depreciation', 'Accumulated Depr', 'Ending BV']
            headers_ok = 0
            for col_idx, expected in enumerate(expected_headers, 1):
                actual = ws_depr.cell(row=1, column=col_idx).value
                if actual is not None and str(actual).strip().lower() == expected.lower():
                    headers_ok += 1

            # Check first data row (row 2)
            a2 = ws_depr.cell(row=2, column=1).value  # Year = 1
            b2 = ws_depr.cell(row=2, column=2).value  # Beginning BV = asset cost (a number)
            c2 = ws_depr.cell(row=2, column=3).value  # Depreciation = reference to annual depr
            d2 = ws_depr.cell(row=2, column=4).value  # Accumulated Depr = C2 or similar
            e2 = ws_depr.cell(row=2, column=5).value  # Ending BV = B2-C2

            # Check year 1 marker
            year1_ok = (a2 == 1)
            # Check beginning BV is a number (the asset cost) for year 1
            bv_ok = isinstance(b2, (int, float)) and b2 > 0
            # Check depreciation references Assets sheet
            depr_ok = isinstance(c2, str) and 'ASSETS' in c2.upper() and 'E' in c2.upper()
            # Check ending BV formula
            ending_bv_ok = isinstance(e2, str) and 'B2' in e2.upper() and 'C2' in e2.upper()

            # Check that data rows exist (more than just headers)
            data_rows = ws_depr.max_row - 1  # subtract header row

            if headers_ok >= 4 and year1_ok and bv_ok and depr_ok and data_rows >= 5:
                print(f"PASS: Component 3 — DepreciationSchedule has {headers_ok}/5 headers, year sequence starts at 1, {data_rows} data rows, depreciation references Assets sheet")
                total_score += 0.25
            elif headers_ok >= 3 and data_rows >= 3:
                print(f"PARTIAL: Component 3 — DepreciationSchedule partially filled: headers={headers_ok}/5, data_rows={data_rows}, year1_ok={year1_ok}")
                total_score += 0.12
            else:
                print(f"FAIL: Component 3 — DepreciationSchedule insufficient: headers={headers_ok}/5, data_rows={data_rows}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ---- Component 4: Stacked bar chart in DepreciationSchedule (0.25 points) ----
    # Task asks for stacked bar chart showing net book value vs accumulated depreciation
    # Initial sheet has no chart; golden adds a stacked bar chart with 2 series
    try:
        if 'DepreciationSchedule' not in wb.sheetnames:
            print("FAIL: Component 4 — 'DepreciationSchedule' sheet not found")
        else:
            ws_depr = wb['DepreciationSchedule']
            charts = ws_depr._charts

            if len(charts) == 0:
                print("FAIL: Component 4 — No chart found in DepreciationSchedule")
            else:
                chart = charts[0]
                chart_type = type(chart).__name__
                # Check it's a bar chart
                is_bar = 'Bar' in chart_type
                # Check it's stacked
                is_stacked = hasattr(chart, 'grouping') and chart.grouping in ('stacked', 'percentStacked')
                # Check it has 2 series (NBV and Accumulated Depr)
                series_count = len(chart.series)
                has_two_series = series_count >= 2

                if is_bar and is_stacked and has_two_series:
                    print(f"PASS: Component 4 — Stacked bar chart with {series_count} series found in DepreciationSchedule (type={chart_type}, grouping={chart.grouping})")
                    total_score += 0.25
                elif is_bar and has_two_series:
                    print(f"PARTIAL: Component 4 — Bar chart with {series_count} series found, but not stacked (grouping={getattr(chart, 'grouping', 'N/A')})")
                    total_score += 0.12
                elif len(charts) >= 1:
                    print(f"PARTIAL: Component 4 — Chart found but wrong type or series count (type={chart_type}, series={series_count}, stacked={is_stacked})")
                    total_score += 0.08
                else:
                    print(f"FAIL: Component 4 — Chart is not a stacked bar chart with 2 series (type={chart_type}, stacked={is_stacked}, series={series_count})")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
