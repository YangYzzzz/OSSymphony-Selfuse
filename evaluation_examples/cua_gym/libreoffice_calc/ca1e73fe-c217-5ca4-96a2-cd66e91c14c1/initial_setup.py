"""
Initial Setup: Inventory management spreadsheet with raw data.
Task ID: calc_ops_043
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_043'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Inventory ---
    ws = wb.active
    ws.title = 'Inventory'

    headers = ['SKU', 'Description', 'Qty', 'Unit Cost', 'Reorder Point', 'Daily Usage']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    data = [
        ['SK-01', 'Widget A',  500,  12.00, 200, 25],
        ['SK-02', 'Widget B',  150,   8.50, 300, 40],
        ['SK-03', 'Gadget C', 1200,   3.00, 400, 50],
        ['SK-04', 'Gadget D',   80,  45.00, 100, 10],
        ['SK-05', 'Part E',   2000,   1.50, 500, 80],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
