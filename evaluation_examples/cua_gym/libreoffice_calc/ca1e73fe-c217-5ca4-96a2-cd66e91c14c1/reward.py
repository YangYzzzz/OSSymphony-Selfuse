"""
Reward Script: Inventory management dashboard with formulas
Task ID: calc_ops_043
Domain: libreoffice_calc
Scoring:
  C1: Dashboard sheet exists (0.15)
  C2: Dashboard title in A1 (0.10)
  C3: Total SKUs formula in B3 — COUNTA (0.20)
  C4: Total Inventory Value formula in B4 — SUMPRODUCT (0.20)
  C5: Items Below Reorder formula in B5 — count where C<E (0.20)
  C6: Avg Days of Supply formula in B6 — AVERAGE of C/F (0.15)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_043'


def normalize_formula(f):
    """Normalize a formula string for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Dashboard sheet exists (0.15 points)
    try:
        sheet_names_lower = [s.lower() for s in wb.sheetnames]
        if 'dashboard' in sheet_names_lower:
            # Get the actual sheet name with original casing
            dash_idx = sheet_names_lower.index('dashboard')
            dash_name = wb.sheetnames[dash_idx]
            ws = wb[dash_name]
            print(f"PASS: Component 1 — Dashboard sheet exists as '{dash_name}' (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — No 'Dashboard' sheet found. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0  # No dashboard means nothing else to check
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 2: Dashboard title in A1 (0.10 points)
    try:
        a1_val = ws['A1'].value
        if a1_val and 'inventory' in str(a1_val).lower() and 'dashboard' in str(a1_val).lower():
            print(f"PASS: Component 2 — A1 contains title: '{a1_val}' (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 — Expected 'Inventory Dashboard' in A1, found: {repr(a1_val)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Total SKUs formula in B3 using COUNTA (0.20 points)
    # Ground truth: =COUNTA(Inventory.A2:A6) = 5
    try:
        b3_val = ws['B3'].value
        b3_norm = normalize_formula(b3_val)
        if 'COUNTA(' in b3_norm or 'COUNT(' in b3_norm:
            # Check it references the Inventory sheet A column range
            if 'A2' in b3_norm and 'A6' in b3_norm:
                print(f"PASS: Component 3 — B3 has count formula referencing Inventory A2:A6: '{b3_val}' (0.20 pts)")
                total_score += 0.20
            else:
                # Still accept if it's a COUNTA covering the right range somehow
                print(f"PARTIAL: Component 3 — B3 has count formula but range unclear: '{b3_val}' (0.10 pts)")
                total_score += 0.10
        elif isinstance(b3_val, (int, float)) and b3_val == 5:
            # Hardcoded correct value — partial credit
            print(f"PARTIAL: Component 3 — B3 has correct value 5 but no formula (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — Expected COUNTA formula in B3, found: {repr(b3_val)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Total Inventory Value formula in B4 using SUMPRODUCT (0.20 points)
    # Ground truth: =SUMPRODUCT(Inventory.C2:C6,Inventory.D2:D6) = 16330
    try:
        b4_val = ws['B4'].value
        b4_norm = normalize_formula(b4_val)
        if 'SUMPRODUCT(' in b4_norm:
            # Check it references Qty (C) and Unit Cost (D) columns
            if ('C2' in b4_norm and 'D2' in b4_norm):
                print(f"PASS: Component 4 — B4 has SUMPRODUCT formula with C and D columns: '{b4_val}' (0.20 pts)")
                total_score += 0.20
            else:
                print(f"PARTIAL: Component 4 — B4 has SUMPRODUCT but columns unclear: '{b4_val}' (0.10 pts)")
                total_score += 0.10
        elif isinstance(b4_val, (int, float)) and abs(b4_val - 16330) < 1:
            print(f"PARTIAL: Component 4 — B4 has correct value 16330 but no formula (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 — Expected SUMPRODUCT formula in B4, found: {repr(b4_val)}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Items Below Reorder formula in B5 (0.20 points)
    # Ground truth: count where C<E = 2 (SK-02 and SK-04)
    # Could be COUNTIF, SUMPRODUCT with comparison, etc.
    try:
        b5_val = ws['B5'].value
        b5_norm = normalize_formula(b5_val)
        if isinstance(b5_val, str) and b5_val.startswith('='):
            # Accept any formula that compares C (Qty) vs E (Reorder Point)
            has_c_col = bool(re.search(r'C\d', b5_norm))
            has_e_col = bool(re.search(r'E\d', b5_norm))
            if has_c_col and has_e_col:
                print(f"PASS: Component 5 — B5 has formula comparing Qty vs Reorder: '{b5_val}' (0.20 pts)")
                total_score += 0.20
            else:
                print(f"PARTIAL: Component 5 — B5 has a formula but may not compare C vs E: '{b5_val}' (0.10 pts)")
                total_score += 0.10
        elif isinstance(b5_val, (int, float)) and b5_val == 2:
            print(f"PARTIAL: Component 5 — B5 has correct value 2 but no formula (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 — Expected formula for items below reorder in B5, found: {repr(b5_val)}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Avg Days of Supply formula in B6 (0.15 points)
    # Ground truth: =AVERAGE(Inventory.C2:C6/Inventory.F2:F6) = 16.15
    # Days of supply = Qty / Daily Usage, then average
    try:
        b6_val = ws['B6'].value
        b6_norm = normalize_formula(b6_val)
        if isinstance(b6_val, str) and b6_val.startswith('='):
            # Check for AVERAGE and references to C (Qty) and F (Daily Usage)
            has_avg = 'AVERAGE(' in b6_norm
            has_c_col = bool(re.search(r'C\d', b6_norm))
            has_f_col = bool(re.search(r'F\d', b6_norm))
            if has_avg and has_c_col and has_f_col:
                print(f"PASS: Component 6 — B6 has AVERAGE formula with Qty/DailyUsage: '{b6_val}' (0.15 pts)")
                total_score += 0.15
            elif has_c_col and has_f_col:
                print(f"PARTIAL: Component 6 — B6 references C and F but not AVERAGE: '{b6_val}' (0.07 pts)")
                total_score += 0.07
            else:
                print(f"PARTIAL: Component 6 — B6 has formula but doesn't reference expected cols: '{b6_val}' (0.05 pts)")
                total_score += 0.05
        elif isinstance(b6_val, (int, float)) and abs(b6_val - 16.15) < 0.5:
            print(f"PARTIAL: Component 6 — B6 has approx correct value {b6_val} but no formula (0.07 pts)")
            total_score += 0.07
        else:
            print(f"FAIL: Component 6 — Expected AVERAGE formula in B6, found: {repr(b6_val)}")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
