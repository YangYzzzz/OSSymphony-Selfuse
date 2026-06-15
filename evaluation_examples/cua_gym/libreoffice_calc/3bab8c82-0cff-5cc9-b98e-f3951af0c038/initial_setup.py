"""
Initial Setup: Create a Revenue spreadsheet with monthly data and a lookup area.
Task ID: calc_lf_023
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_023'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Revenue ---
    ws = wb.active
    ws.title = 'Revenue'

    # Headers
    ws['A1'] = 'Month'
    ws['B1'] = 'Amount'
    ws['D1'] = 'Start Month'
    ws['E1'] = 'Sum of 3 Months'

    # Style headers
    header_font = Font(bold=True, size=11, name='Calibri')
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    for cell_ref in ['A1', 'B1', 'D1', 'E1']:
        cell = ws[cell_ref]
        cell.font = Font(bold=True, size=11, name='Calibri', color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align

    # Monthly data
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun']
    amounts = [10000, 12000, 11500, 13000, 14500, 15000]

    for i, (month, amount) in enumerate(zip(months, amounts), start=2):
        ws.cell(row=i, column=1, value=month)
        ws.cell(row=i, column=2, value=amount)
        ws.cell(row=i, column=2).number_format = '#,##0'

    # Lookup area
    ws['D2'] = 'Mar'
    # E2 intentionally left EMPTY - this is where the agent must enter the formula

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 4
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
