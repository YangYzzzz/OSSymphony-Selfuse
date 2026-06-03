"""
Reward Script: Create an OFFSET-based dynamic range that starts at the matched row
               and spans 3 rows down, then SUM those values.
Task ID: calc_lf_023
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): E2 contains a formula (not empty/literal)
  Component 2 (0.3): Formula uses both OFFSET and MATCH functions
  Component 3 (0.3): Formula structure is correct - SUM(OFFSET(B1,MATCH(D2,...),0,3,1))
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_023'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Revenue sheet exists
    if 'Revenue' not in wb.sheetnames:
        print("FAIL: 'Revenue' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Revenue']

    # Get E2 value
    e2_val = ws['E2'].value

    # Component 1: E2 contains a formula (0.4 points)
    # This differentiates initial (E2 is None) from golden (E2 has formula)
    try:
        if e2_val is not None and isinstance(e2_val, str) and e2_val.startswith('='):
            print(f"PASS: Component 1 - E2 contains a formula: {e2_val} (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 - E2 should contain a formula, found: {repr(e2_val)}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Formula uses OFFSET and MATCH (0.3 points)
    # The task specifically requires an OFFSET-based dynamic range with MATCH
    try:
        if e2_val is not None and isinstance(e2_val, str):
            formula_upper = e2_val.upper().replace(' ', '')
            has_offset = 'OFFSET(' in formula_upper
            has_match = 'MATCH(' in formula_upper
            if has_offset and has_match:
                print(f"PASS: Component 2 - Formula uses both OFFSET and MATCH (0.3 pts)")
                total_score += 0.3
            else:
                missing = []
                if not has_offset:
                    missing.append('OFFSET')
                if not has_match:
                    missing.append('MATCH')
                print(f"FAIL: Component 2 - Formula missing: {', '.join(missing)}")
        else:
            print(f"FAIL: Component 2 - E2 is not a formula string")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Formula structure is correct (0.3 points)
    # Expected: =SUM(OFFSET(B1,MATCH(D2,A2:A7,0),0,3,1))
    # Key structural elements:
    #   - SUM wraps the OFFSET
    #   - OFFSET references column B (the Amount column)
    #   - MATCH references D2 (Start Month) against month range in column A
    #   - OFFSET height parameter is 3 (span 3 rows)
    try:
        if e2_val is not None and isinstance(e2_val, str):
            formula_upper = e2_val.upper().replace(' ', '')

            # Check SUM wraps OFFSET
            has_sum_offset = 'SUM(OFFSET(' in formula_upper

            # Check OFFSET height is 3 (the "3 rows down" requirement)
            # Pattern: OFFSET(...,MATCH(...),<col_offset>,3,<width>)
            # After MATCH closes, we expect ,0,3,1 or similar with height=3
            has_height_3 = bool(re.search(r'OFFSET\([^)]*MATCH\([^)]*\)[^)]*,3,', formula_upper))

            # Check MATCH references D2 (the start month cell)
            has_match_d2 = bool(re.search(r'MATCH\(D2,', formula_upper))

            if has_sum_offset and has_height_3 and has_match_d2:
                print(f"PASS: Component 3 - Formula structure correct: SUM(OFFSET(...MATCH(D2,...)...,3,...)) (0.3 pts)")
                total_score += 0.3
            else:
                details = []
                if not has_sum_offset:
                    details.append('missing SUM(OFFSET(...))')
                if not has_height_3:
                    details.append('OFFSET height not 3')
                if not has_match_d2:
                    details.append('MATCH does not reference D2')
                print(f"FAIL: Component 3 - Structural issues: {'; '.join(details)}")
        else:
            print(f"FAIL: Component 3 - E2 is not a formula string")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
