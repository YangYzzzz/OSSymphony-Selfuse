"""
Initial Setup: Track customer onboarding completion rates for new accounts
Task ID: calc_sales_customer_onboarding_072
Domain: libreoffice_calc
"""

import openpyxl
from datetime import date, timedelta
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_customer_onboarding_072'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'OnboardingTracker'

    # --- Headers ---
    headers = ['Account', 'Contract Date', 'Milestone 1', 'M2', 'M3', 'M4', 'M5', 'M6',
               'Completion %', 'Days Since Start', 'Status']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # --- Realistic account names ---
    accounts = [
        'Apex Dynamics Corp', 'BlueSky Technologies', 'ClearWave Solutions', 'DataNest Inc',
        'Elevate Software', 'FuturePath Systems', 'GlobalEdge Partners', 'HighPoint Analytics',
        'Innovix Networks', 'JetStream Commerce', 'KeyLogic Enterprises', 'Luminary Digital',
        'Meridian Group', 'NorthStar Consulting', 'Orbit Cloud Services', 'PrimeRoute Ltd',
        'QuantumLeap Tech', 'RapidScale AI', 'Skyline Integration', 'TechBridge Global',
        'UltraVision Inc', 'VectorPath Software', 'WaveCrest Media', 'XcelRate Systems',
        'YieldMax Analytics', 'ZenithCore Solutions', 'Alpha Collective', 'BrightWave Corp',
        'CoreAxis Technologies', 'DeepField Networks', 'EaglePoint Ventures', 'FlowForce Digital',
        'GridMind Systems', 'HorizonShift Inc', 'ImpactFirst Tech', 'JumpStart Analytics',
        'KineticLoop Ltd', 'LiftOff Software', 'MindBridge AI', 'NexGen Platforms',
        'OpenLoop Inc', 'PeakPerform Corp', 'QuikSync Technologies', 'RealEdge Partners',
        'Scalar Dynamics', 'TrueNorth Systems', 'UniFront Digital', 'Vantage360 Inc',
        'WideArc Solutions', 'XStream Commerce', 'YellowBrick Consulting', 'ZeroFriction Tech',
        'ArcLight Networks', 'BoldPath Enterprises', 'ClearSign Analytics', 'DriftWave Media',
        'EvolveCore Inc', 'FrontLine Systems', 'GreenSpark Digital', 'HighGear Software',
        'IntelliSync Corp', 'JetPoint Ventures', 'KoreLogic Partners', 'LaunchPad AI',
        'MetaFlow Inc', 'NetBridge Technologies', 'OceanView Systems', 'PulseCore Digital',
        'QuestNet Solutions', 'RoamFree Platforms', 'SkyBridge Corp', 'TitanPath Analytics',
        'UltraEdge Inc', 'VeloCore Technologies', 'WarpSpeed Systems', 'XenonData Corp',
        'YieldPath Digital', 'ZenScale Inc', 'AgileMind Solutions', 'BrightForce Tech',
    ]

    # Contract dates: 15 to 120 days ago (varied to create both "at risk" and "on track" accounts)
    random.seed(42)
    base_date = date(2025, 11, 1)  # approximate "today" - 120 days

    milestone_sets = [
        # (0 done) completely new
        ['', '', '', '', '', ''],
        # (1 done)
        ['Done', '', '', '', '', ''],
        # (2 done)
        ['Done', 'Done', '', '', '', ''],
        # (3 done) 50%
        ['Done', 'Done', 'Done', '', '', ''],
        # (4 done)
        ['Done', 'Done', 'Done', 'Done', '', ''],
        # (5 done)
        ['Done', 'Done', 'Done', 'Done', 'Done', ''],
        # (6 done) complete
        ['Done', 'Done', 'Done', 'Done', 'Done', 'Done'],
    ]

    for i, account in enumerate(accounts):
        row = i + 2
        # Vary contract dates to create interesting distribution
        days_ago = random.randint(5, 100)
        contract_date = base_date - timedelta(days=days_ago - days_ago % 1)
        # Vary completion levels
        milestone_idx = random.randint(0, 6)
        milestones = milestone_sets[milestone_idx]

        ws.cell(row=row, column=1, value=account)
        ws.cell(row=row, column=2, value=contract_date)
        ws.cell(row=row, column=2).number_format = 'yyyy-mm-dd'
        for m, ms_val in enumerate(milestones, 3):
            ws.cell(row=row, column=m, value=ms_val if ms_val else None)
        # Columns I (9), J (10), K (11) left empty — task is to add formulas there

    # Set column widths for readability
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 14
    for col in ['C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 12
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 16
    ws.column_dimensions['K'].width = 12

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

create_initial()
