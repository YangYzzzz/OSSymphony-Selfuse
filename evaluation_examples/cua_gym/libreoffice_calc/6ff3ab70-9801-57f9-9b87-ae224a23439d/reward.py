"""
Reward Script: Track customer onboarding completion rates
Task ID: calc_sales_customer_onboarding_072
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.30): COUNTIF completion % formulas in I2:I81
  - Component 2 (0.20): TODAY()-B# days-since-start formulas in J2:J81
  - Component 3 (0.20): IFS status formulas in K2:K81
  - Component 4 (0.10): Conditional formatting on K2:K81
  - Component 5 (0.10): Bucket summary data in M1:N5 (labels + COUNTIFS formulas)
  - Component 6 (0.10): Bar/column chart present on OnboardingTracker sheet
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_sales_customer_onboarding_072'


def normalize_formula(f):
    """Normalize formula for loose comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Gate: sheet must exist
    if 'OnboardingTracker' not in wb.sheetnames:
        print("CRITICAL: Sheet 'OnboardingTracker' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['OnboardingTracker']

    # Component 1: COUNTIF completion % formulas in I2:I81 (0.30 points)
    # Each formula should be =COUNTIF(CX:HX,"Done")/6
    # We verify all 80 rows have the correct pattern
    try:
        formula_i_correct = 0
        formula_i_total = 80
        for row in range(2, 82):
            val = ws.cell(row=row, column=9).value  # column I
            if isinstance(val, str):
                norm = normalize_formula(val)
                expected = normalize_formula(f'=COUNTIF(C{row}:H{row},"Done")/6')
                if norm == expected:
                    formula_i_correct += 1

        if formula_i_correct == formula_i_total:
            print(f"PASS: Component 1 — All {formula_i_total} COUNTIF formulas in I2:I81 correct (0.30 pts)")
            total_score += 0.30
        elif formula_i_correct >= formula_i_total * 0.9:
            print(f"PARTIAL: Component 1 — {formula_i_correct}/{formula_i_total} COUNTIF formulas correct (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — Only {formula_i_correct}/{formula_i_total} COUNTIF formulas in I2:I81 correct; "
                  f"expected =COUNTIF(Cx:Hx,\"Done\")/6")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: TODAY()-B# days-since-start formulas in J2:J81 (0.20 points)
    try:
        formula_j_correct = 0
        formula_j_total = 80
        for row in range(2, 82):
            val = ws.cell(row=row, column=10).value  # column J
            if isinstance(val, str):
                norm = normalize_formula(val)
                expected = normalize_formula(f'=TODAY()-B{row}')
                if norm == expected:
                    formula_j_correct += 1

        if formula_j_correct == formula_j_total:
            print(f"PASS: Component 2 — All {formula_j_total} TODAY()-Bx formulas in J2:J81 correct (0.20 pts)")
            total_score += 0.20
        elif formula_j_correct >= formula_j_total * 0.9:
            print(f"PARTIAL: Component 2 — {formula_j_correct}/{formula_j_total} TODAY()-Bx formulas correct (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 — Only {formula_j_correct}/{formula_j_total} TODAY()-Bx formulas in J2:J81 correct; "
                  f"expected =TODAY()-Bx")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: IFS status formulas in K2:K81 (0.20 points)
    # Expected pattern: =IFS(Ix=1,"Complete",AND(Ix<0.5,Jx>30),"At Risk",TRUE,"On Track")
    try:
        formula_k_correct = 0
        formula_k_total = 80
        # Pattern: starts with =IFS( and includes "Complete", "At Risk", "On Track"
        ifs_pattern = re.compile(
            r'=IFS\(.*"COMPLETE".*"ATRISK".*"ONTRACK"\)|=IFS\(.*"ATRISK".*"COMPLETE".*"ONTRACK"\)',
            re.IGNORECASE
        )
        for row in range(2, 82):
            val = ws.cell(row=row, column=11).value  # column K
            if isinstance(val, str):
                norm = normalize_formula(val)
                # Check the formula contains IFS with all three status values
                if ('IFS(' in norm and
                        '"COMPLETE"' in norm and
                        '"ATRISK"' in norm and
                        '"ONTRACK"' in norm and
                        f'I{row}' in norm and
                        f'J{row}' in norm):
                    formula_k_correct += 1

        if formula_k_correct == formula_k_total:
            print(f"PASS: Component 3 — All {formula_k_total} IFS status formulas in K2:K81 correct (0.20 pts)")
            total_score += 0.20
        elif formula_k_correct >= formula_k_total * 0.9:
            print(f"PARTIAL: Component 3 — {formula_k_correct}/{formula_k_total} IFS formulas correct (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — Only {formula_k_correct}/{formula_k_total} IFS formulas in K2:K81 correct; "
                  f"expected IFS with Complete/At Risk/On Track based on I and J columns")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Conditional formatting on K2:K81 column (0.10 points)
    # Should have at least one CF rule covering K2:K81 with expression rules
    try:
        cf_found = False
        cf_count = 0
        for cf in ws.conditional_formatting:
            cf_str = str(cf)
            # Check if this CF range covers K column
            if 'K' in cf_str:
                rule_count = len(cf.rules)
                if rule_count >= 1:
                    cf_found = True
                    cf_count = rule_count

        if cf_found:
            print(f"PASS: Component 4 — Conditional formatting found on K column "
                  f"({cf_count} rules) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 — No conditional formatting found on K column (K2:K81); "
                  f"expected rules for At Risk/On Track/Complete")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Bucket summary in M/N columns (0.10 points)
    # M1='Completion Bucket', N1='Count'
    # M2:M5 = '0-25%', '26-50%', '51-75%', '76-100%'
    # N2:N5 = COUNTIFS formulas
    try:
        bucket_score = 0.0
        # Check headers
        m1 = ws.cell(row=1, column=13).value
        n1 = ws.cell(row=1, column=14).value
        headers_ok = (m1 is not None and n1 is not None)

        # Check bucket labels
        expected_labels = ['0-25%', '26-50%', '51-75%', '76-100%']
        labels_found = 0
        for i, label in enumerate(expected_labels, 2):
            actual = ws.cell(row=i, column=13).value
            if actual and str(actual).strip() == label:
                labels_found += 1

        # Check COUNTIFS formulas in N2:N5
        countifs_found = 0
        for row in range(2, 6):
            val = ws.cell(row=row, column=14).value
            if isinstance(val, str) and 'COUNTIFS' in val.upper():
                countifs_found += 1

        if headers_ok and labels_found == 4 and countifs_found == 4:
            print(f"PASS: Component 5 — Bucket summary complete: "
                  f"headers, 4 labels, 4 COUNTIFS formulas (0.10 pts)")
            total_score += 0.10
        elif labels_found >= 2 or countifs_found >= 2:
            print(f"PARTIAL: Component 5 — Partial bucket summary: "
                  f"labels={labels_found}/4, COUNTIFS={countifs_found}/4 (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — Bucket summary incomplete: "
                  f"headers={headers_ok}, labels={labels_found}/4, COUNTIFS={countifs_found}/4; "
                  f"expected M1='Completion Bucket', N1='Count', M2:M5 labels, N2:N5 COUNTIFS")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Bar/column chart present on the sheet (0.10 points)
    try:
        charts = ws._charts
        chart_found = False
        chart_type_desc = ''
        for chart in charts:
            chart_type_desc = type(chart).__name__
            # BarChart covers both bar and column types
            if 'Bar' in chart_type_desc or 'bar' in chart_type_desc.lower():
                chart_found = True
                break

        if not chart_found and len(charts) > 0:
            # Accept any chart type as partial compliance
            chart_found = True
            chart_type_desc = type(charts[0]).__name__

        if chart_found:
            print(f"PASS: Component 6 — Chart found on OnboardingTracker: "
                  f"type={chart_type_desc} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 6 — No chart found on OnboardingTracker sheet; "
                  f"expected bar/column chart showing completion rate distribution")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
