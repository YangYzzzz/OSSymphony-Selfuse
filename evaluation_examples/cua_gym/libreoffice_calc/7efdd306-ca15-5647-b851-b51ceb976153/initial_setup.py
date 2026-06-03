"""
Initial Setup: TEXTJOIN+IF formula task - concatenate active user names
Task ID: calc_fma_textjoin_if_053
Domain: libreoffice_calc
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fma_textjoin_if_053'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: ActiveUsers ---
    ws = wb.active
    ws.title = 'ActiveUsers'

    # Headers
    ws['A1'] = 'Name'
    ws['B1'] = 'Status'
    ws['D1'] = 'Active Users List:'

    # 15 users with roughly alternating Active/Inactive (9 Active, 6 Inactive)
    # Active on rows: 2, 3, 5, 7, 8, 10, 12, 14, 15  (9 Active)
    # Inactive on rows: 4, 6, 9, 11, 13, 16            (6 Inactive)
    users = [
        ('Sarah Chen', 'Active'),
        ('Marcus Johnson', 'Active'),
        ('Diana Patel', 'Inactive'),
        ('Ethan Brooks', 'Active'),
        ('Fiona Larsen', 'Inactive'),
        ('George Nakamura', 'Active'),
        ('Hannah Rivera', 'Active'),
        ('Ivan Okonkwo', 'Inactive'),
        ('Julia Santana', 'Active'),
        ('Kevin Müller', 'Inactive'),
        ('Laura Kim', 'Active'),
        ('Noah Fitzgerald', 'Inactive'),
        ('Olivia Hartmann', 'Active'),
        ('Patrick O\'Brien', 'Inactive'),
        ('Qiana Washington', 'Active'),
    ]

    for i, (name, status) in enumerate(users, 2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=status)

    # D2 is intentionally empty (task requires placing formula there)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['D'].width = 50

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
