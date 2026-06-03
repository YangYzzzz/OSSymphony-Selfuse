"""
Reward Script: Concatenate active user names using TEXTJOIN+IF formula in cell D2
Task ID: calc_fma_textjoin_if_053
Domain: libreoffice_calc
Scoring:
  Component 1: D2 is not empty — has a formula or value (0.3 pts)
  Component 2: D2 formula uses TEXTJOIN and IF filtering on column B for 'Active' (0.4 pts)
  Component 3: D2 formula covers the correct full data range B2:B16 / A2:A16 (0.3 pts)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fma_textjoin_if_053'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task: Place a single-cell formula in D2 that concatenates all names from
    column A (rows 2-16) where the corresponding status in column B is 'Active',
    separated by commas. The expected formula:
      =TEXTJOIN(", ",TRUE,IF(B2:B16="Active",A2:A16,""))
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'ActiveUsers' sheet must exist
    if 'ActiveUsers' not in wb.sheetnames:
        print("FAIL: Sheet 'ActiveUsers' not found in workbook")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws = wb['ActiveUsers']

    # Component 1: D2 is not empty — contains a formula or non-null value (0.3 points)
    # In the initial file, D2 is None. The task requires placing a formula there.
    try:
        d2_value = ws['D2'].value
        if d2_value is not None and str(d2_value).strip() != '':
            print(f"PASS: Component 1 — D2 is not empty (value: {repr(str(d2_value)[:80])}) (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — D2 is empty (None or blank); task requires a formula here")
    except Exception as e:
        print(f"ERROR: Component 1 — Cannot read D2: {e}")

    # Component 2: D2 formula uses TEXTJOIN with IF filtering 'Active' from column B (0.4 points)
    # The formula must contain TEXTJOIN and IF with "Active" condition on column B
    try:
        d2_value = ws['D2'].value
        if d2_value is None:
            print(f"FAIL: Component 2 — D2 is empty, cannot check TEXTJOIN+IF formula")
        else:
            formula_str = str(d2_value).strip().upper().replace(' ', '')
            # Must be a formula (starts with =)
            is_formula = formula_str.startswith('=')
            # Must contain TEXTJOIN
            has_textjoin = 'TEXTJOIN' in formula_str
            # Must contain IF
            has_if = 'IF(' in formula_str
            # Must filter on "ACTIVE" in column B
            has_active_condition = '"ACTIVE"' in formula_str or "'ACTIVE'" in formula_str
            has_col_b_condition = re.search(r'B\d+:B\d+.*ACTIVE|ACTIVE.*B\d+:B\d+', formula_str) is not None or \
                                   ('B2:B16' in formula_str.replace(' ', '') and 'ACTIVE' in formula_str)

            if is_formula and has_textjoin and has_if and has_active_condition:
                print(f"PASS: Component 2 — D2 contains TEXTJOIN+IF formula filtering 'Active' values (0.4 pts)")
                print(f"  Formula: {repr(str(d2_value)[:100])}")
                total_score += 0.4
            else:
                issues = []
                if not is_formula:
                    issues.append("does not start with '='")
                if not has_textjoin:
                    issues.append("missing TEXTJOIN function")
                if not has_if:
                    issues.append("missing IF function")
                if not has_active_condition:
                    issues.append("missing 'Active' condition string")
                print(f"FAIL: Component 2 — Formula issues: {'; '.join(issues)}")
                print(f"  Found in D2: {repr(str(d2_value)[:100])}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: D2 formula references the correct full data range B2:B16 and A2:A16 (0.3 points)
    # The task specifies names in A2:A16 and statuses in B2:B16 (15 rows of data)
    try:
        d2_value = ws['D2'].value
        if d2_value is None:
            print(f"FAIL: Component 3 — D2 is empty, cannot check range coverage")
        else:
            formula_str = str(d2_value).strip().replace(' ', '')
            formula_upper = formula_str.upper()
            # Must reference column A range covering rows 2-16
            has_a_range = bool(re.search(r'A2:A1[6-9]|A2:A[2-9]\d', formula_upper))
            # Must reference column B range covering rows 2-16
            has_b_range = bool(re.search(r'B2:B1[6-9]|B2:B[2-9]\d', formula_upper))

            # More permissive: check at least B2:B16 and A2:A16 literally
            has_a16 = 'A2:A16' in formula_upper
            has_b16 = 'B2:B16' in formula_upper

            if (has_a_range or has_a16) and (has_b_range or has_b16):
                print(f"PASS: Component 3 — Formula covers correct full data range A2:A16 and B2:B16 (0.3 pts)")
                total_score += 0.3
            else:
                missing = []
                if not (has_a_range or has_a16):
                    missing.append("A column range A2:A16")
                if not (has_b_range or has_b16):
                    missing.append("B column range B2:B16")
                print(f"FAIL: Component 3 — Formula missing correct range reference: {', '.join(missing)}")
                print(f"  Found in D2: {repr(str(d2_value)[:100])}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
