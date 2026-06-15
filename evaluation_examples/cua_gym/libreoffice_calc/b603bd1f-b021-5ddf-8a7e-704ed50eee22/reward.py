"""
Reward Script: Freeze top two rows in LibreOffice Calc spreadsheet
Task ID: calc_gfl_044
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): freeze_panes is exactly 'A3' (rows 1-2 frozen)
  Component 2 (0.3): freeze is set AND merged title row A1:I1 is intact
  Component 3 (0.2): freeze is set AND header row 2 has correct column names
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_044'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice document."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify that freeze panes are set to A3 (top two rows frozen).
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active
    if ws is None:
        print("CRITICAL: No active sheet found")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: freeze_panes is exactly 'A3' (0.5 points)
    # This is the core task requirement: freeze rows 1-2
    try:
        freeze_val = ws.freeze_panes
        if freeze_val == 'A3':
            print(f"PASS: Component 1 -- freeze_panes is 'A3' (0.5 pts)")
            total_score += 0.5
        elif freeze_val is not None:
            # Partial: some freeze is set but not the correct one
            print(f"FAIL: Component 1 -- freeze_panes is '{freeze_val}', expected 'A3'")
        else:
            print(f"FAIL: Component 1 -- freeze_panes is None, expected 'A3'")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: freeze is set AND merged title row A1:I1 is intact (0.3 points)
    # Compound check: freeze must be correct + title row preserved
    try:
        freeze_ok = (ws.freeze_panes == 'A3')
        title_val = ws['A1'].value
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        title_merged = any('A1' in r and 'I1' in r for r in merged_ranges)

        if freeze_ok and title_val == 'Customer Transaction Log 2024' and title_merged:
            print(f"PASS: Component 2 -- freeze set AND title row intact: '{title_val}', merged A1:I1 (0.3 pts)")
            total_score += 0.3
        elif not freeze_ok:
            print(f"FAIL: Component 2 -- freeze not correctly set (prerequisite for this component)")
        else:
            print(f"FAIL: Component 2 -- title='{title_val}', merged={title_merged}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: freeze is set AND header row 2 has correct column names (0.2 points)
    # Compound check: freeze must be correct + headers preserved
    try:
        freeze_ok = (ws.freeze_panes == 'A3')
        expected_headers = [
            'Transaction ID', 'Date', 'Customer', 'Product',
            'Qty', 'Unit Price', 'Total', 'Payment Method', 'Status'
        ]
        actual_headers = []
        for col in range(1, 10):
            val = ws.cell(row=2, column=col).value
            actual_headers.append(val)

        headers_match = (actual_headers == expected_headers)

        if freeze_ok and headers_match:
            print(f"PASS: Component 3 -- freeze set AND headers intact: {actual_headers} (0.2 pts)")
            total_score += 0.2
        elif not freeze_ok:
            print(f"FAIL: Component 3 -- freeze not correctly set (prerequisite for this component)")
        else:
            print(f"FAIL: Component 3 -- headers mismatch: {actual_headers}")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
