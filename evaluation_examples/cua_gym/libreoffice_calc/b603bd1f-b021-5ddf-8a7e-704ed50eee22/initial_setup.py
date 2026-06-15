"""
Initial Setup: Freeze top two rows in a spreadsheet
Task ID: calc_gfl_044
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_044'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # --- Row 1: Merged title ---
    ws.merge_cells("A1:I1")
    ws["A1"] = "Customer Transaction Log 2024"
    ws["A1"].font = Font(name="Arial", size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # --- Row 2: Headers ---
    headers = [
        "Transaction ID", "Date", "Customer", "Product", "Qty",
        "Unit Price", "Total", "Payment Method", "Status"
    ]
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Data generation ---
    first_names = [
        "Sarah", "Marcus", "Emily", "David", "Rachel", "James", "Olivia",
        "Michael", "Sophia", "Daniel", "Ava", "Christopher", "Isabella",
        "Matthew", "Mia", "Andrew", "Charlotte", "Joshua", "Amelia", "Ryan",
        "Harper", "Nathan", "Evelyn", "Brandon", "Abigail", "Tyler", "Emma",
        "Kevin", "Grace", "Justin", "Lily", "Aaron", "Chloe", "Derek",
        "Ella", "Sean", "Aria", "Kyle", "Nora", "Brian",
    ]
    last_names = [
        "Chen", "Johnson", "Williams", "Garcia", "Martinez", "Anderson",
        "Taylor", "Thomas", "Moore", "Jackson", "Martin", "Lee", "Thompson",
        "White", "Harris", "Clark", "Lewis", "Robinson", "Walker", "Young",
        "Allen", "King", "Wright", "Scott", "Green", "Baker", "Adams",
        "Nelson", "Hill", "Ramirez", "Campbell", "Mitchell", "Roberts",
        "Carter", "Phillips", "Evans", "Turner", "Torres", "Parker", "Collins",
    ]
    products = [
        ("Wireless Mouse", 29.99), ("USB-C Hub", 45.50), ("Mechanical Keyboard", 89.99),
        ("Monitor Stand", 34.95), ("Webcam HD", 59.99), ("Laptop Sleeve", 24.50),
        ("Desk Lamp", 42.00), ("Headset Pro", 79.99), ("Cable Organizer", 12.99),
        ("Mousepad XL", 19.95), ("Phone Charger", 15.99), ("Screen Protector", 9.99),
        ("Bluetooth Speaker", 54.99), ("Power Strip", 22.50), ("Notebook Stand", 38.99),
        ("Drawing Tablet", 149.99), ("External SSD 1TB", 89.00), ("USB Flash Drive 64GB", 14.99),
        ("HDMI Cable 6ft", 11.99), ("Surge Protector", 27.50),
    ]
    payment_methods = ["Credit Card", "Debit Card", "PayPal", "Bank Transfer", "Cash"]
    statuses = ["Completed", "Pending", "Processing", "Shipped", "Delivered", "Refunded"]

    # Generate 198 transaction rows (rows 3-200)
    for i in range(198):
        row = i + 3
        tid = f"TXN-2024-{10001 + i}"
        # Random date in 2024
        month = random.randint(1, 12)
        day = random.randint(1, 28)
        date_str = f"2024-{month:02d}-{day:02d}"
        customer = f"{random.choice(first_names)} {random.choice(last_names)}"
        product_name, unit_price = random.choice(products)
        qty = random.randint(1, 10)
        total = round(qty * unit_price, 2)
        payment = random.choice(payment_methods)
        status = random.choice(statuses)

        ws.cell(row=row, column=1, value=tid)
        ws.cell(row=row, column=2, value=date_str)
        ws.cell(row=row, column=3, value=customer)
        ws.cell(row=row, column=4, value=product_name)
        ws.cell(row=row, column=5, value=qty)
        ws.cell(row=row, column=6, value=unit_price)
        ws.cell(row=row, column=6).number_format = '$#,##0.00'
        ws.cell(row=row, column=7, value=total)
        ws.cell(row=row, column=7).number_format = '$#,##0.00'
        ws.cell(row=row, column=8, value=payment)
        ws.cell(row=row, column=9, value=status)

    # Set column widths for readability
    col_widths = {"A": 18, "B": 14, "C": 22, "D": 22, "E": 8, "F": 14, "G": 14, "H": 18, "I": 14}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # NO freeze panes - the task is to add them
    ws.freeze_panes = None

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
