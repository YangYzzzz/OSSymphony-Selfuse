"""
Reward Script: Citation analysis for Turing Award ML winners
Task ID: osworld_multi_apps_web_scholar_012
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): File 'turing_award_ml.ods' exists on Desktop
  Component 2 (0.30): File contains all 9 required columns
  Component 3 (0.20): File contains Trend_Analysis column with non-empty content
  Component 4 (0.20): Data rows present with valid ML/AI Turing Award data
                       (at least 3 rows, Year in 2018-2023, numeric H_Index / Total_Citations)
Total: 1.0
"""

import os
import shutil
import tempfile
import zipfile

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_web_scholar_012'
ODS_PATH = '/home/user/Desktop/turing_award_ml.ods'

REQUIRED_COLUMNS = [
    'Year',
    'Recipient(s)',
    'Award_Reason',
    'H_Index',
    'Total_Citations',
    'Most_Cited_Paper',
    'Citation_Count_Most_Cited',
    'Career_Span_Years',
    'Avg_Papers_Per_Year',
]

TREND_COLUMN = 'Trend_Analysis'


def load_spreadsheet(file_path):
    """
    Load the spreadsheet file. The .ods file may actually be an XLSX zip
    (i.e. saved with a wrong extension). We detect this using zipfile and
    copy it to a .xlsx temp file so openpyxl can load it without invoking external commands.

    Returns an openpyxl Workbook, or None on failure.
    """
    import openpyxl

    tmp_dir = tempfile.mkdtemp()

    # Check if the file is actually an XLSX-format zip (very common in this environment)
    if zipfile.is_zipfile(file_path):
        with zipfile.ZipFile(file_path) as z:
            names = z.namelist()
        # XLSX files contain xl/ directory structure; ODS files contain content.xml
        if any(n.startswith('xl/') for n in names):
            # It is an XLSX stored with .ods extension — copy and reload
            xlsx_copy = os.path.join(tmp_dir, 'turing_award_ml.xlsx')
            shutil.copy(file_path, xlsx_copy)
            try:
                wb = openpyxl.load_workbook(xlsx_copy)
                print("INFO: Loaded ODS file as XLSX (xlsx structure detected inside .ods container)")
                shutil.rmtree(tmp_dir, ignore_errors=True)
                return wb
            except Exception as e:
                print(f"WARN: Could not load xlsx copy: {e}")
        else:
            # Genuine ODS — try pandas as fallback
            try:
                import pandas as pd
                xl_copy = os.path.join(tmp_dir, 'turing_award_ml_pd.ods')
                shutil.copy(file_path, xl_copy)
                df = pd.read_excel(xl_copy, engine='odf')
                # Convert to a simple in-memory openpyxl workbook so the rest of
                # the verification logic is uniform
                wb = openpyxl.Workbook()
                ws = wb.active
                # Write headers
                for ci, col in enumerate(df.columns, 1):
                    ws.cell(row=1, column=ci, value=col)
                # Write data
                for ri, row in enumerate(df.itertuples(index=False), 2):
                    for ci, val in enumerate(row, 1):
                        ws.cell(row=ri, column=ci, value=val)
                print("INFO: Loaded genuine ODS via pandas/odf engine")
                shutil.rmtree(tmp_dir, ignore_errors=True)
                return wb
            except Exception as e:
                print(f"WARN: pandas ODF load failed: {e}")

    # Last resort: try openpyxl directly (may work if extension is wrong)
    try:
        wb = openpyxl.load_workbook(file_path)
        shutil.rmtree(tmp_dir, ignore_errors=True)
        return wb
    except Exception as e:
        print(f"WARN: Direct openpyxl load failed: {e}")

    shutil.rmtree(tmp_dir, ignore_errors=True)
    return None


def verify_task():
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # ------------------------------------------------------------------ #
    # Component 1: File 'turing_award_ml.ods' exists on the Desktop (0.30 points)
    # This FAILS on initial_env (file not present) and PASSES on golden_env.
    # ------------------------------------------------------------------ #
    try:
        if os.path.exists(ODS_PATH):
            print(f"PASS: Component 1 — turing_award_ml.ods exists on Desktop (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 — turing_award_ml.ods not found at {ODS_PATH}")
            # Without the file there is nothing else to verify
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # ------------------------------------------------------------------ #
    # Load spreadsheet for further inspection
    # ------------------------------------------------------------------ #
    wb = load_spreadsheet(ODS_PATH)
    if wb is None:
        print("CRITICAL: Cannot load spreadsheet — aborting further checks")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Use the first (active) sheet
    ws = wb.active

    # ------------------------------------------------------------------ #
    # Component 2: All 9 required column headers are present (0.30 points)
    # Checks that the agent populated the header row with the expected labels.
    # This FAILS on initial_env (no file) and PASSES on golden_env.
    # ------------------------------------------------------------------ #
    try:
        # Read row-1 header values
        actual_headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        # Normalise: strip whitespace
        actual_headers_norm = [str(h).strip() if h is not None else '' for h in actual_headers]

        missing_cols = []
        for req_col in REQUIRED_COLUMNS:
            # Case-insensitive match
            if not any(req_col.lower() == h.lower() for h in actual_headers_norm):
                missing_cols.append(req_col)

        if not missing_cols:
            print(f"PASS: Component 2 — All 9 required columns present: {actual_headers_norm} (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 2 — Missing columns: {missing_cols}. Found: {actual_headers_norm}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ------------------------------------------------------------------ #
    # Component 3: Trend_Analysis column exists and has non-empty content (0.20 points)
    # This verifies the "trend analysis column" requirement from the task.
    # This FAILS on initial_env (no file) and PASSES on golden_env.
    # ------------------------------------------------------------------ #
    try:
        actual_headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        actual_headers_norm = [str(h).strip() if h is not None else '' for h in actual_headers]

        trend_col_idx = None
        for idx, h in enumerate(actual_headers_norm, 1):
            if h.lower() == TREND_COLUMN.lower():
                trend_col_idx = idx
                break

        if trend_col_idx is None:
            print(f"FAIL: Component 3 — Trend_Analysis column not found in headers: {actual_headers_norm}")
        else:
            # Check that at least one data row has non-empty trend content
            trend_values = []
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=trend_col_idx).value
                if val and str(val).strip():
                    trend_values.append(str(val).strip())

            if trend_values:
                print(f"PASS: Component 3 — Trend_Analysis column has {len(trend_values)} non-empty row(s). "
                      f"First value: {trend_values[0][:80]!r} (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 3 — Trend_Analysis column exists at col {trend_col_idx} "
                      f"but all values are empty")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ------------------------------------------------------------------ #
    # Component 4: Data rows contain valid ML/AI Turing Award data (0.20 points)
    # Checks:
    #   - At least 3 data rows (min from 2018 winners: Bengio, Hinton, LeCun)
    #   - Year values are in range 2018-2023
    #   - H_Index values are numeric (> 0)
    #   - Total_Citations values are numeric (> 0)
    # This FAILS on initial_env (no file) and PASSES on golden_env.
    # ------------------------------------------------------------------ #
    try:
        actual_headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        actual_headers_norm = [str(h).strip() if h is not None else '' for h in actual_headers]

        def find_col(col_name):
            for idx, h in enumerate(actual_headers_norm, 1):
                if h.lower() == col_name.lower():
                    return idx
            return None

        year_col = find_col('Year')
        h_index_col = find_col('H_Index')
        citations_col = find_col('Total_Citations')

        data_row_count = ws.max_row - 1  # subtract header row
        if data_row_count < 3:
            print(f"FAIL: Component 4 — Only {data_row_count} data row(s) found; expected at least 3")
        elif year_col is None or h_index_col is None or citations_col is None:
            print(f"FAIL: Component 4 — Required columns for data check not found "
                  f"(Year:{year_col}, H_Index:{h_index_col}, Total_Citations:{citations_col})")
        else:
            valid_rows = 0
            invalid_details = []
            for row in range(2, ws.max_row + 1):
                year_val = ws.cell(row=row, column=year_col).value
                h_idx_val = ws.cell(row=row, column=h_index_col).value
                cit_val = ws.cell(row=row, column=citations_col).value

                def _year_valid(v):
                    try:
                        return 2018 <= int(v) <= 2023
                    except (TypeError, ValueError):
                        return False

                def _positive_numeric(v):
                    try:
                        return float(v) > 0
                    except (TypeError, ValueError):
                        return False

                year_ok = _year_valid(year_val)
                h_ok = _positive_numeric(h_idx_val)
                cit_ok = _positive_numeric(cit_val)

                if year_ok and h_ok and cit_ok:
                    valid_rows += 1
                else:
                    invalid_details.append(
                        f"Row {row}: year={year_val}(ok={year_ok}), "
                        f"h_index={h_idx_val}(ok={h_ok}), "
                        f"citations={cit_val}(ok={cit_ok})"
                    )

            if valid_rows >= 3:
                print(f"PASS: Component 4 — {valid_rows} valid data rows with Year 2018-2023, "
                      f"numeric H_Index and Total_Citations (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 4 — Only {valid_rows} valid data rows. "
                      f"Issues: {invalid_details[:3]}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


if __name__ == '__main__':
    verify_task()
