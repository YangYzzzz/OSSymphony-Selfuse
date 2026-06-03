"""
Initial Setup: Create annual_reviews.xlsx and review_template.pdf for multi-app annual review PDF task
Task ID: osworld_multi_apps_excel_pdf_form_012
Domain: libreoffice_calc (multi-app: also uses PDF template)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_excel_pdf_form_012'
EXCEL_OUTPUT = f'{WORKDIR}/annual_reviews.xlsx'
TEMPLATE_OUTPUT = f'{WORKDIR}/Desktop/review_template.pdf'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_excel():
    """Create annual_reviews.xlsx with employee performance data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Annual Reviews"

    # Header styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Headers
    headers = [
        "EmployeeID", "Name", "Department",
        "Technical Skills (1-5)", "Communication (1-5)",
        "Leadership (1-5)", "Teamwork (1-5)",
        "Overall Score", "Manager Recommendation"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # Set column widths
    col_widths = [12, 22, 16, 22, 20, 18, 16, 14, 22]
    col_letters = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
    for letter, width in zip(col_letters, col_widths):
        ws.column_dimensions[letter].width = width
    ws.row_dimensions[1].height = 36

    # Realistic employee data: EmployeeID, Name, Department, Technical(1-5),
    # Communication(1-5), Leadership(1-5), Teamwork(1-5), Overall, Recommendation
    employees = [
        ["EMP001", "Sarah Chen",        "Engineering",   5, 4, 4, 5, 4.5,  "Promote"],
        ["EMP002", "Marcus Johnson",    "Engineering",   4, 5, 3, 4, 4.0,  "Retain"],
        ["EMP003", "Priya Patel",       "Marketing",     3, 5, 4, 5, 4.25, "Retain"],
        ["EMP004", "Derek Williams",    "Marketing",     2, 3, 2, 3, 2.5,  "PIP"],
        ["EMP005", "Aisha Okonkwo",     "HR",            4, 5, 5, 5, 4.75, "Promote"],
        ["EMP006", "James Kowalski",    "Engineering",   5, 3, 4, 4, 4.0,  "Retain"],
        ["EMP007", "Lisa Nguyen",       "Finance",       4, 4, 3, 4, 3.75, "Retain"],
        ["EMP008", "Roberto Martinez",  "Finance",       3, 3, 2, 3, 2.75, "PIP"],
        ["EMP009", "Fatima Al-Hassan",  "HR",            5, 5, 5, 5, 5.0,  "Promote"],
        ["EMP010", "Tyler Brooks",      "Marketing",     4, 4, 4, 4, 4.0,  "Retain"],
        ["EMP011", "Hannah Kimura",     "Engineering",   5, 4, 5, 5, 4.75, "Promote"],
        ["EMP012", "David Osei",        "Finance",       4, 5, 4, 4, 4.25, "Retain"],
    ]

    data_align = Alignment(horizontal="center", vertical="center")
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, emp in enumerate(employees, 2):
        ws.row_dimensions[row_idx].height = 18
        for col_idx, val in enumerate(emp, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = data_align
            cell.border = data_border
            # Color recommendation column
            if col_idx == 9:
                if val == "Promote":
                    cell.font = Font(color="FF006100")
                    cell.fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
                elif val == "PIP":
                    cell.font = Font(color="FF9C0006")
                    cell.fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(EXCEL_OUTPUT)
    print(f'Excel file created: {EXCEL_OUTPUT}')


def create_pdf_template():
    """Create review_template.pdf on the Desktop."""
    from fpdf import FPDF

    os.makedirs(f'{WORKDIR}/Desktop', exist_ok=True)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_margins(15, 15, 15)

    # Title
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(47, 84, 150)
    pdf.cell(0, 12, "Annual Performance Review", ln=True, align="C")
    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 7, "Employee Annual Review Form", ln=True, align="C")
    pdf.ln(4)

    # Employee Info section
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_fill_color(47, 84, 150)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 8, "  Employee Information", ln=True, fill=True)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 11)
    pdf.ln(2)
    pdf.cell(40, 7, "Employee ID:", ln=False)
    pdf.cell(0, 7, "___________________", ln=True)
    pdf.cell(40, 7, "Name:", ln=False)
    pdf.cell(0, 7, "___________________", ln=True)
    pdf.cell(40, 7, "Department:", ln=False)
    pdf.cell(0, 7, "___________________", ln=True)
    pdf.cell(40, 7, "Review Year:", ln=False)
    pdf.cell(0, 7, "2024", ln=True)
    pdf.ln(4)

    # Competency Ratings section
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_fill_color(47, 84, 150)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 8, "  Competency Ratings", ln=True, fill=True)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(2)

    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, "Rate each competency on a scale of 1 (Needs Improvement) to 5 (Outstanding).", ln=True)
    pdf.ln(2)

    competencies = [
        "Technical Skills",
        "Communication",
        "Leadership",
        "Teamwork",
    ]

    pdf.set_font("Helvetica", "B", 11)
    for comp in competencies:
        pdf.cell(55, 8, comp + ":", ln=False)
        # Draw 5 rating boxes with labels
        for i in range(1, 6):
            pdf.cell(6, 8, "[ ]", ln=False)
            pdf.cell(5, 8, str(i), ln=False)
            if i < 5:
                pdf.cell(3, 8, "", ln=False)
        pdf.ln(8)

    pdf.ln(4)

    # Overall Score section
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_fill_color(47, 84, 150)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 8, "  Overall Score", ln=True, fill=True)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(2)
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(55, 8, "Overall Score (average):", ln=False)
    pdf.cell(0, 8, "___________", ln=True)
    pdf.ln(4)

    # Manager Recommendation section
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_fill_color(47, 84, 150)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 8, "  Manager Recommendation", ln=True, fill=True)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(2)
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 7, "Select one:", ln=True)
    pdf.ln(1)
    pdf.cell(8, 7, "[ ]", ln=False)
    pdf.cell(0, 7, "Promote   - Exceptional performance, ready for advancement", ln=True)
    pdf.cell(8, 7, "[ ]", ln=False)
    pdf.cell(0, 7, "Retain    - Meeting expectations, continue in current role", ln=True)
    pdf.cell(8, 7, "[ ]", ln=False)
    pdf.cell(0, 7, "PIP       - Performance Improvement Plan required", ln=True)
    pdf.ln(6)

    # Signature section
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_fill_color(47, 84, 150)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 8, "  Signatures", ln=True, fill=True)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(2)
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(90, 7, "Manager Signature: _______________", ln=False)
    pdf.cell(0, 7, "Date: _______________", ln=True)
    pdf.cell(90, 7, "Employee Signature: _______________", ln=False)
    pdf.cell(0, 7, "Date: _______________", ln=True)

    pdf.output(TEMPLATE_OUTPUT)
    print(f'PDF template created: {TEMPLATE_OUTPUT}')


def create_initial():
    create_excel()
    create_pdf_template()

    # Ensure reviews output folder does NOT exist (it should be created by the agent)
    reviews_dir = f'{WORKDIR}/Desktop/reviews'
    if os.path.exists(reviews_dir):
        import shutil
        shutil.rmtree(reviews_dir)
        print(f'Removed pre-existing reviews directory: {reviews_dir}')

    # GUI-ready startup: open annual_reviews.xlsx in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{EXCEL_OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with annual_reviews.xlsx (DISPLAY=:0)')


create_initial()
