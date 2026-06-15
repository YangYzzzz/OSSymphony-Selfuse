"""
Initial Setup: Delete the 'Scratch' sheet from the workbook
Task ID: calc_gg1_038
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_038'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # Shared styles
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    def style_header_row(ws, headers, row=1):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    # --- Sheet 1: Dashboard ---
    ws_dash = wb.active
    ws_dash.title = "Dashboard"

    ws_dash["A1"] = "Regional Sales Dashboard - FY2025"
    ws_dash["A1"].font = Font(name="Calibri", size=16, bold=True, color="2F5496")
    ws_dash["A2"] = "Last updated: 2025-06-30"
    ws_dash["A2"].font = Font(name="Calibri", size=10, italic=True, color="808080")

    dash_headers = ["Region", "Q1 Revenue", "Q2 Revenue", "Total", "Growth %"]
    style_header_row(ws_dash, dash_headers, row=4)

    dash_data = [
        ["Northeast", 284500, 312800, 597300, 9.9],
        ["Southeast", 198700, 215400, 414100, 8.4],
        ["Midwest", 167300, 178900, 346200, 6.9],
        ["Southwest", 143200, 156700, 299900, 9.4],
        ["West Coast", 321600, 348200, 669800, 8.3],
        ["Pacific NW", 112400, 125600, 238000, 11.7],
    ]
    for r, row_data in enumerate(dash_data, 5):
        for c, val in enumerate(row_data, 1):
            cell = ws_dash.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c in (2, 3, 4):
                cell.number_format = '$#,##0'
            elif c == 5:
                cell.number_format = '0.0%'

    ws_dash.column_dimensions["A"].width = 16
    ws_dash.column_dimensions["B"].width = 15
    ws_dash.column_dimensions["C"].width = 15
    ws_dash.column_dimensions["D"].width = 15
    ws_dash.column_dimensions["E"].width = 12

    # --- Sheet 2: Q1 Data ---
    ws_q1 = wb.create_sheet("Q1 Data")
    q1_headers = ["Date", "Sales Rep", "Client", "Product", "Units", "Unit Price", "Total Amount", "Status"]
    style_header_row(ws_q1, q1_headers)

    q1_data = [
        ["2025-01-08", "Sarah Chen", "Apex Manufacturing", "Enterprise Suite", 3, 12500, 37500, "Closed"],
        ["2025-01-12", "Marcus Johnson", "BrightPath Labs", "Standard License", 10, 3200, 32000, "Closed"],
        ["2025-01-19", "Elena Rodriguez", "Coastal Dynamics", "Premium Add-on", 5, 7800, 39000, "Closed"],
        ["2025-02-03", "David Kim", "DataFlow Inc", "Enterprise Suite", 2, 12500, 25000, "Closed"],
        ["2025-02-11", "Sarah Chen", "EverGreen Solutions", "Standard License", 15, 3200, 48000, "Closed"],
        ["2025-02-18", "Marcus Johnson", "FrostByte Systems", "Premium Add-on", 8, 7800, 62400, "Pending"],
        ["2025-02-25", "Elena Rodriguez", "GlobalTech Corp", "Enterprise Suite", 1, 12500, 12500, "Closed"],
        ["2025-03-05", "David Kim", "Highland Analytics", "Standard License", 20, 3200, 64000, "Closed"],
        ["2025-03-14", "Sarah Chen", "InnoVate Partners", "Premium Add-on", 4, 7800, 31200, "Closed"],
        ["2025-03-22", "Marcus Johnson", "JetStream Media", "Enterprise Suite", 2, 12500, 25000, "Pending"],
        ["2025-03-28", "Elena Rodriguez", "KnowledgeBase AI", "Standard License", 12, 3200, 38400, "Closed"],
        ["2025-03-31", "David Kim", "LunarTech Solutions", "Premium Add-on", 6, 7800, 46800, "Closed"],
    ]
    for r, row_data in enumerate(q1_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_q1.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 6:
                cell.number_format = '$#,##0'
            elif c == 7:
                cell.number_format = '$#,##0'

    for col_letter, width in [("A", 12), ("B", 18), ("C", 22), ("D", 18), ("E", 8), ("F", 12), ("G", 14), ("H", 10)]:
        ws_q1.column_dimensions[col_letter].width = width

    ws_q1.freeze_panes = "A2"

    # --- Sheet 3: Q2 Data ---
    ws_q2 = wb.create_sheet("Q2 Data")
    style_header_row(ws_q2, q1_headers)

    q2_data = [
        ["2025-04-02", "Sarah Chen", "MountainView Tech", "Enterprise Suite", 4, 12500, 50000, "Closed"],
        ["2025-04-10", "Marcus Johnson", "NexGen Robotics", "Standard License", 8, 3200, 25600, "Closed"],
        ["2025-04-18", "Elena Rodriguez", "OceanBlue Pharma", "Premium Add-on", 7, 7800, 54600, "Closed"],
        ["2025-05-01", "David Kim", "PrimePath Consulting", "Enterprise Suite", 3, 12500, 37500, "Pending"],
        ["2025-05-09", "Sarah Chen", "QuantumLeap AI", "Standard License", 18, 3200, 57600, "Closed"],
        ["2025-05-15", "Marcus Johnson", "RedStone Industries", "Premium Add-on", 5, 7800, 39000, "Closed"],
        ["2025-05-23", "Elena Rodriguez", "SilverLake Data", "Enterprise Suite", 2, 12500, 25000, "Closed"],
        ["2025-06-04", "David Kim", "TerraFirm Analytics", "Standard License", 14, 3200, 44800, "Closed"],
        ["2025-06-12", "Sarah Chen", "UltraVision Corp", "Premium Add-on", 9, 7800, 70200, "Closed"],
        ["2025-06-19", "Marcus Johnson", "VentureForge LLC", "Enterprise Suite", 1, 12500, 12500, "Closed"],
        ["2025-06-25", "Elena Rodriguez", "WaveRider Systems", "Standard License", 11, 3200, 35200, "Pending"],
        ["2025-06-30", "David Kim", "XenoTech Labs", "Premium Add-on", 3, 7800, 23400, "Closed"],
    ]
    for r, row_data in enumerate(q2_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_q2.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 6:
                cell.number_format = '$#,##0'
            elif c == 7:
                cell.number_format = '$#,##0'

    for col_letter, width in [("A", 12), ("B", 18), ("C", 22), ("D", 18), ("E", 8), ("F", 12), ("G", 14), ("H", 10)]:
        ws_q2.column_dimensions[col_letter].width = width

    ws_q2.freeze_panes = "A2"

    # --- Sheet 4: Charts ---
    ws_charts = wb.create_sheet("Charts")
    ws_charts["A1"] = "Chart Data Summary"
    ws_charts["A1"].font = Font(name="Calibri", size=14, bold=True)

    chart_headers = ["Quarter", "Enterprise Suite", "Standard License", "Premium Add-on"]
    style_header_row(ws_charts, chart_headers, row=3)

    chart_data = [
        ["Q1 2025", 100000, 182400, 179400],
        ["Q2 2025", 125000, 163200, 187200],
    ]
    for r, row_data in enumerate(chart_data, 4):
        for c, val in enumerate(row_data, 1):
            cell = ws_charts.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c >= 2:
                cell.number_format = '$#,##0'

    for col_letter, width in [("A", 12), ("B", 18), ("C", 18), ("D", 18)]:
        ws_charts.column_dimensions[col_letter].width = width

    # --- Sheet 5: Scratch (to be deleted by agent) ---
    ws_scratch = wb.create_sheet("Scratch")
    ws_scratch["A1"] = "DRAFT - Temporary Calculations"
    ws_scratch["A1"].font = Font(name="Calibri", size=12, bold=True, color="FF0000")

    ws_scratch["A3"] = "Test formula check"
    ws_scratch["B3"] = 45230
    ws_scratch["C3"] = 38900
    ws_scratch["D3"] = "=B3+C3"

    ws_scratch["A5"] = "Rough revenue estimate"
    ws_scratch["B5"] = 1200000
    ws_scratch["A6"] = "Adjusted target"
    ws_scratch["B6"] = 1350000
    ws_scratch["A7"] = "Variance"
    ws_scratch["B7"] = "=B6-B5"

    ws_scratch["A9"] = "Notes: need to verify Q2 numbers with finance team"
    ws_scratch["A10"] = "TODO: clean up before sharing"
    ws_scratch["A10"].font = Font(color="FF0000", italic=True)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
