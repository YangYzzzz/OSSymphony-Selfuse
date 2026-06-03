"""
Initial Setup: Decimal separator settings task - spreadsheet with US number formats
Task ID: osworld_calc_decimal_separator_004
Domain: libreoffice_calc

Creates a spreadsheet with realistic financial/business data using English/US number formats.
The agent must change LibreOffice locale settings to German-style number formatting
(comma as decimal separator, period as thousands separator).
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_decimal_separator_004'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Sales Report ---
    ws1 = wb.active
    ws1.title = "Sales Report"

    # Header row styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2E75B6", end_color="FF2E75B6", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Headers
    headers = ["Product", "Region", "Q1 Sales", "Q2 Sales", "Q3 Sales", "Q4 Sales", "Annual Total", "Growth %"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Realistic sales data with US-style decimal values
    data = [
        ["Laptop Pro X",      "North America", 128450.50, 134920.75, 142300.25, 158600.00, 564271.50,  8.45],
        ["Wireless Headset",  "North America",  45230.00,  52180.50,  48900.75,  61450.25, 207761.50, 12.30],
        ["Smart Monitor 4K",  "Europe",          98760.25, 103420.00, 115680.50, 122340.75, 440201.50,  7.85],
        ["Mechanical Keyboard","Europe",          32100.75,  35640.50,  38920.25,  41350.00, 148011.50,  9.20],
        ["USB-C Hub Pro",     "Asia Pacific",    23450.00,  27830.75,  31240.50,  35680.25, 118201.50, 14.75],
        ["Webcam HD 1080p",   "Asia Pacific",    18920.50,  22340.25,  19870.75,  24560.00,  85691.50,  6.50],
        ["Gaming Mouse Elite","North America",   41230.75,  45680.50,  43920.25,  52340.00, 183171.50, 10.15],
        ["Portable SSD 1TB",  "Europe",          67890.25,  71340.50,  78920.75,  84560.00, 302711.50,  8.90],
        ["Tablet Stand Pro",  "Asia Pacific",    12340.50,  14520.75,  16890.25,  19230.00,  62981.50, 13.45],
        ["Noise Cancel Earbuds","North America",  29870.75,  33450.25,  37820.50,  42340.75, 143482.25, 11.60],
        ["Docking Station",   "Europe",          54320.00,  58670.75,  63240.50,  71890.25, 248121.50,  9.75],
        ["Smart Presenter",   "Asia Pacific",     8920.25,  10340.50,  12180.75,  14560.00,  46001.50, 15.20],
    ]

    data_font = Font(name="Calibri", size=11)
    data_align_center = Alignment(horizontal="center")
    data_align_right = Alignment(horizontal="right")

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = border
            if c in (1, 2):
                cell.alignment = data_align_center
            elif isinstance(val, float):
                cell.alignment = data_align_right
                if c in (3, 4, 5, 6, 7):
                    # US-style number format: period as decimal, comma as thousands
                    cell.number_format = '#,##0.00'
                elif c == 8:
                    cell.number_format = '0.00"%"'

    # Column widths
    col_widths = [22, 15, 14, 14, 14, 14, 14, 10]
    for col_idx, width in enumerate(col_widths, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Freeze header row
    ws1.freeze_panes = "A2"

    # --- Sheet 2: Regional Summary ---
    ws2 = wb.create_sheet("Regional Summary")

    # Headers for summary sheet
    summary_headers = ["Region", "Total Revenue", "Avg Deal Size", "Market Share %", "YoY Growth %"]
    for col, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Summary data
    summary_data = [
        ["North America", 1098686.75, 274671.69, 41.25, 10.63],
        ["Europe",        1138046.00, 284511.50, 42.72,  8.90],
        ["Asia Pacific",   312876.00,  78219.00, 11.73, 13.47],
        ["Total",         2549608.75, 212467.40, 95.70, 11.00],
    ]

    for r, row_data in enumerate(summary_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = border
            if c == 1:
                cell.alignment = data_align_center
            elif isinstance(val, float):
                cell.alignment = data_align_right
                if c in (2, 3):
                    cell.number_format = '#,##0.00'
                elif c in (4, 5):
                    cell.number_format = '0.00"%"'

    # Make "Total" row bold
    for c in range(1, 6):
        ws2.cell(row=5, column=c).font = Font(name="Calibri", size=11, bold=True)

    summary_col_widths = [18, 16, 16, 16, 14]
    for col_idx, width in enumerate(summary_col_widths, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws2.freeze_panes = "A2"

    # --- Sheet 3: Exchange Rates ---
    ws3 = wb.create_sheet("Exchange Rates")

    ex_headers = ["Currency", "Code", "Rate to USD", "Inverse Rate", "Last Updated"]
    for col, h in enumerate(ex_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    exchange_data = [
        ["Euro",            "EUR", 0.92340, 1.08300, "2025-03-01"],
        ["British Pound",   "GBP", 0.78920, 1.26720, "2025-03-01"],
        ["Japanese Yen",    "JPY", 149.850, 0.00668, "2025-03-01"],
        ["Canadian Dollar", "CAD", 1.35420, 0.73840, "2025-03-01"],
        ["Swiss Franc",     "CHF", 0.89640, 1.11560, "2025-03-01"],
        ["Australian Dollar","AUD", 1.52370, 0.65630, "2025-03-01"],
        ["Chinese Yuan",    "CNY", 7.23450, 0.13823, "2025-03-01"],
    ]

    for r, row_data in enumerate(exchange_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = border
            if c in (1, 2, 5):
                cell.alignment = data_align_center
            elif isinstance(val, float):
                cell.alignment = data_align_right
                cell.number_format = '#,##0.00000'

    ex_col_widths = [18, 8, 14, 14, 14]
    for col_idx, width in enumerate(ex_col_widths, 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws3.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the spreadsheet in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
