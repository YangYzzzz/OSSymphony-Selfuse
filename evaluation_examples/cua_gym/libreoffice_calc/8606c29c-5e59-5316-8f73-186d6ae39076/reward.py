"""
Reward Script: Change number format settings to German-style decimal/thousands separators
Task ID: osworld_calc_decimal_separator_004
Domain: libreoffice_calc

Task: Change the number format settings in this spreadsheet to use a comma as the
decimal separator and a period as the thousands separator (German locale style).

Scoring Rubric:
  Component 1: All numeric cells in 'Sales Report' use German locale format [$-407]  — 0.5 points
  Component 2: All numeric cells in 'Regional Summary' use German locale format       — 0.3 points
  Component 3: All numeric cells in 'Exchange Rates' use German locale format         — 0.2 points
  Total: 1.0

German locale format characteristics:
  - Format code contains '[$-407]' prefix (German locale identifier 0x407 = de-DE)
  - Uses comma ',' as decimal separator in format string (e.g., #.##0,00)
  - Uses period '.' as thousands separator in format string
  Examples: '[$-407]#.##0,00', '[$-407]0,00"%"', '[$-407]#.##0,00000'

Initial state:
  - Cells use English/US locale format: '#,##0.00', '0.00"%"', '#,##0.00000'
  - No [$-407] prefix present anywhere

Golden state:
  - ALL numeric-formatted cells use [$-407] German locale format codes
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_decimal_separator_004'


def is_german_locale_format(number_format: str) -> bool:
    """
    Check if a number format code uses German locale ([$-407]) prefix.
    German locale uses:
      - '[$-407]' prefix to declare de-DE locale
      - comma ',' as decimal separator
      - period '.' as thousands separator
    """
    if not number_format or number_format == 'General':
        return False
    # Check for the German locale identifier [$-407]
    return '-407' in number_format


def check_sheet_german_formats(wb, sheet_name):
    """
    Check how many numeric-formatted cells in the given sheet use German locale.
    Returns (german_count, total_count, non_german_formats) tuple.
    """
    ws = wb[sheet_name]
    german_count = 0
    total_count = 0
    non_german_formats = set()

    for row in ws.iter_rows():
        for cell in row:
            nf = cell.number_format
            if nf and nf != 'General':
                total_count += 1
                if is_german_locale_format(nf):
                    german_count += 1
                else:
                    non_german_formats.add(nf)

    return german_count, total_count, non_german_formats


def verify_task(file_path):
    """
    Verify task completion: all numeric format cells should use German locale
    number format codes ([$-407] prefix with comma decimal, period thousands).
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify all expected sheets exist (precondition gate)
    expected_sheets = ['Sales Report', 'Regional Summary', 'Exchange Rates']
    for sheet_name in expected_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"CRITICAL: Expected sheet '{sheet_name}' not found. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0

    # Component 1: 'Sales Report' sheet — all numeric cells use German locale format (0.5 points)
    # The main sales data sheet has the most cells — most important to verify
    try:
        german_count, total_count, non_german = check_sheet_german_formats(wb, 'Sales Report')
        if total_count == 0:
            print("FAIL: Component 1 — 'Sales Report' has no numeric-formatted cells (unexpected)")
        elif german_count == total_count:
            print(f"PASS: Component 1 — 'Sales Report': all {total_count} numeric cells use German locale format [$-407] (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — 'Sales Report': {german_count}/{total_count} cells use German locale format.")
            if non_german:
                print(f"       Non-German formats still present: {non_german}")
    except Exception as e:
        print(f"ERROR: Component 1 — 'Sales Report' check failed: {e}")

    # Component 2: 'Regional Summary' sheet — all numeric cells use German locale format (0.3 points)
    try:
        german_count, total_count, non_german = check_sheet_german_formats(wb, 'Regional Summary')
        if total_count == 0:
            print("FAIL: Component 2 — 'Regional Summary' has no numeric-formatted cells (unexpected)")
        elif german_count == total_count:
            print(f"PASS: Component 2 — 'Regional Summary': all {total_count} numeric cells use German locale format [$-407] (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — 'Regional Summary': {german_count}/{total_count} cells use German locale format.")
            if non_german:
                print(f"       Non-German formats still present: {non_german}")
    except Exception as e:
        print(f"ERROR: Component 2 — 'Regional Summary' check failed: {e}")

    # Component 3: 'Exchange Rates' sheet — all numeric cells use German locale format (0.2 points)
    try:
        german_count, total_count, non_german = check_sheet_german_formats(wb, 'Exchange Rates')
        if total_count == 0:
            print("FAIL: Component 3 — 'Exchange Rates' has no numeric-formatted cells (unexpected)")
        elif german_count == total_count:
            print(f"PASS: Component 3 — 'Exchange Rates': all {total_count} numeric cells use German locale format [$-407] (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — 'Exchange Rates': {german_count}/{total_count} cells use German locale format.")
            if non_german:
                print(f"       Non-German formats still present: {non_german}")
    except Exception as e:
        print(f"ERROR: Component 3 — 'Exchange Rates' check failed: {e}")

    final_score = round(min(total_score, 1.0), 1)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
