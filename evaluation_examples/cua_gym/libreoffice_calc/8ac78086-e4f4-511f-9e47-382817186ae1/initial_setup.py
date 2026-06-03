"""
Initial Setup: Fix INDIRECT formula with sheet names containing spaces
Task ID: calc_tbl_072
Domain: libreoffice_calc

Sheets named 'Sheet 1', 'Sheet 2', 'Sheet 3' with data.
Dashboard sheet has A1=' 1' and a broken INDIRECT formula in B2
that does not handle spaces in sheet names.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_072'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Regional Sales ---
    ws1 = wb.active
    ws1.title = 'Sheet 1'
    headers1 = ['Region', 'Q1 Revenue', 'Q2 Revenue', 'Q3 Revenue', 'Q4 Revenue', 'Total']
    for col, h in enumerate(headers1, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")

    data1 = [
        ['Northeast', 45230, 52100, 48750, 61200, None],
        ['Southeast', 38900, 41500, 39800, 45600, None],
        ['Midwest', 52100, 48300, 55200, 58900, None],
        ['West Coast', 67800, 71200, 69500, 74300, None],
        ['Mountain', 28400, 31200, 29800, 33100, None],
        ['Pacific NW', 41600, 44800, 42300, 47200, None],
        ['Southwest', 35200, 38700, 36900, 40100, None],
        ['Central', 49300, 52600, 50800, 55400, None],
        ['Atlantic', 56700, 59800, 57200, 62100, None],
        ['Great Lakes', 43200, 46500, 44800, 48900, None],
    ]
    for r, row_data in enumerate(data1, 2):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c_idx, value=val)
            if c_idx >= 2 and c_idx <= 5 and val is not None:
                cell.number_format = '$#,##0'
        # Total formula
        ws1.cell(row=r, column=6, value=f'=SUM(B{r}:E{r})')
        ws1.cell(row=r, column=6).number_format = '$#,##0'

    # Summary row
    ws1.cell(row=12, column=1, value='Grand Total')
    ws1.cell(row=12, column=1).font = Font(bold=True)
    for c_idx in range(2, 7):
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(c_idx)
        ws1.cell(row=12, column=c_idx, value=f'=SUM({col_letter}2:{col_letter}11)')
        ws1.cell(row=12, column=c_idx).number_format = '$#,##0'
        ws1.cell(row=12, column=c_idx).font = Font(bold=True)

    # Set column widths
    ws1.column_dimensions['A'].width = 15
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws1.column_dimensions[col_letter].width = 14

    # --- Sheet 2: Product Categories ---
    ws2 = wb.create_sheet('Sheet 2')
    headers2 = ['Category', 'Units Sold', 'Avg Price', 'Revenue', 'Margin %']
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color="FF70AD47", end_color="FF70AD47", fill_type="solid")
        c.font = Font(bold=True, color="FFFFFF")

    data2 = [
        ['Electronics', 1250, 89.99, 112487.50, 0.22],
        ['Clothing', 3400, 34.50, 117300.00, 0.45],
        ['Home & Garden', 890, 56.75, 50507.50, 0.35],
        ['Sports', 2100, 42.00, 88200.00, 0.30],
        ['Books', 5600, 15.99, 89544.00, 0.55],
        ['Food & Beverage', 8200, 8.50, 69700.00, 0.18],
        ['Automotive', 420, 125.00, 52500.00, 0.25],
        ['Health', 1800, 28.75, 51750.00, 0.40],
        ['Toys', 3100, 19.99, 61969.00, 0.38],
        ['Office Supplies', 4500, 12.25, 55125.00, 0.42],
    ]
    for r, row_data in enumerate(data2, 2):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c_idx, value=val)
            if c_idx == 3 or c_idx == 4:
                cell.number_format = '$#,##0.00'
            elif c_idx == 5:
                cell.number_format = '0%'

    ws2.column_dimensions['A'].width = 18
    for col_letter in ['B', 'C', 'D', 'E']:
        ws2.column_dimensions[col_letter].width = 14

    # --- Sheet 3: Monthly Trends ---
    ws3 = wb.create_sheet('Sheet 3')
    headers3 = ['Month', 'Sales', 'Returns', 'Net Sales', 'Growth %']
    for col, h in enumerate(headers3, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color="FFED7D31", end_color="FFED7D31", fill_type="solid")
        c.font = Font(bold=True, color="FFFFFF")

    data3 = [
        ['January', 82500, 4125, 78375, None],
        ['February', 79300, 3965, 75335, -0.0388],
        ['March', 91200, 4560, 86640, 0.1501],
        ['April', 88700, 4435, 84265, -0.0274],
        ['May', 95400, 4770, 90630, 0.0755],
        ['June', 102300, 5115, 97185, 0.0723],
        ['July', 98600, 4930, 93670, -0.0362],
        ['August', 104500, 5225, 99275, 0.0599],
        ['September', 97800, 4890, 92910, -0.0641],
        ['October', 108200, 5410, 102790, 0.1063],
        ['November', 115600, 5780, 109820, 0.0684],
        ['December', 128400, 6420, 121980, 0.1107],
    ]
    for r, row_data in enumerate(data3, 2):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r, column=c_idx, value=val)
            if c_idx in [2, 3, 4] and val is not None:
                cell.number_format = '$#,##0'
            elif c_idx == 5 and val is not None:
                cell.number_format = '0.00%'

    ws3.column_dimensions['A'].width = 14
    for col_letter in ['B', 'C', 'D', 'E']:
        ws3.column_dimensions[col_letter].width = 14

    # --- Dashboard sheet: has BROKEN INDIRECT formula ---
    ws_dash = wb.create_sheet('Dashboard', 0)  # Insert at position 0

    ws_dash['A1'] = ' 1'  # Leading space + '1' = ' 1'
    ws_dash['A1'].font = Font(size=11)

    # Label cells
    ws_dash['A3'] = 'Dynamic Sheet Reference Tool'
    ws_dash['A3'].font = Font(size=14, bold=True)

    ws_dash['A5'] = 'Sheet Number (in A1):'
    ws_dash['B5'] = '=A1'
    ws_dash['A6'] = 'Target Cell:'
    ws_dash['B6'] = 'B5'

    ws_dash['A8'] = 'Lookup Result:'
    ws_dash['A8'].font = Font(bold=True)

    # BROKEN formula - missing single quotes for sheet names with spaces
    # This produces "Sheet 1.B5" but INDIRECT needs "'Sheet 1'.B5"
    ws_dash['B8'] = '=INDIRECT("Sheet"&A1&".B5")'

    ws_dash['A10'] = 'Instructions:'
    ws_dash['A10'].font = Font(bold=True, italic=True)
    ws_dash['A11'] = 'The formula in B8 tries to pull data from a sheet'
    ws_dash['A12'] = 'whose name is built dynamically from cell A1.'
    ws_dash['A13'] = 'Currently it fails because sheet names contain spaces.'
    ws_dash['A14'] = 'Fix the INDIRECT formula to handle spaces in sheet names.'

    # Additional context rows
    ws_dash['A16'] = 'Expected sheet reference format:'
    ws_dash['B16'] = "'Sheet 1'.B5"
    ws_dash['A17'] = 'Current (broken) reference:'
    ws_dash['B17'] = "Sheet 1.B5"

    ws_dash.column_dimensions['A'].width = 35
    ws_dash.column_dimensions['B'].width = 25

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
