"""
Reward Script: Fix INDIRECT formula for sheet names with spaces
Task ID: calc_tbl_072
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): Formula in B8 contains single-quote wrapping around sheet name
  Component 2 (0.3): Formula still uses INDIRECT with dynamic A1 reference
  Component 3 (0.2): Formula resolves to correct value (67800 from 'Sheet 1'.B5)
"""

import os
import re
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_072'


def persist_app_state(domain):
    """Save any unsaved changes in LibreOffice before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Dashboard sheet exists
    if 'Dashboard' not in wb.sheetnames:
        print("CRITICAL: 'Dashboard' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Dashboard']

    # Get the formula in B8
    b8_value = ws['B8'].value
    print(f"DEBUG: B8 raw value = {repr(b8_value)}")

    # Component 1: Formula contains single-quote wrapping around sheet name (0.5 points)
    # The fix requires adding single quotes (apostrophes) around the dynamically-built
    # sheet name in the INDIRECT formula. The broken formula is:
    #   =INDIRECT("Sheet"&A1&".B5")
    # The fixed formula should be something like:
    #   =INDIRECT("'Sheet"&A1&"'.B5")
    # We check that the formula string contains apostrophe/single-quote characters
    # that wrap the sheet name portion.
    try:
        if b8_value is not None and isinstance(b8_value, str):
            formula_upper = b8_value.upper().replace(" ", "")
            # Check for the presence of single quotes (apostrophes) in the formula
            # that serve to quote the sheet name. The key change is adding ' around
            # the sheet name in the INDIRECT string argument.
            # Valid patterns: ="'Sheet"&A1&"'.B5" or variants with single quotes
            has_single_quotes = "'" in b8_value
            # Also verify it's not just the broken original formula
            # The broken formula: =INDIRECT("Sheet"&A1&".B5")
            broken_pattern = '=INDIRECT("Sheet"&A1&".B5")'
            is_broken = (b8_value.replace(" ", "") == broken_pattern.replace(" ", ""))

            if has_single_quotes and not is_broken:
                print(f"PASS: Component 1 -- Formula contains single-quote wrapping ({b8_value}) (0.5 pts)")
                total_score += 0.5
            else:
                print(f"FAIL: Component 1 -- Formula lacks single-quote fix. Value: {b8_value}")
        else:
            print(f"FAIL: Component 1 -- B8 is not a formula string. Value: {repr(b8_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Formula is fixed AND still dynamic (0.3 points)
    # Must have single-quote fix AND preserve INDIRECT+A1 dynamic reference structure.
    # This component only passes if the formula has both the fix AND the dynamic nature.
    # On initial_env, the formula lacks quotes, so this FAILS.
    try:
        if b8_value is not None and isinstance(b8_value, str):
            formula_upper = b8_value.upper().replace(" ", "")
            has_indirect = "INDIRECT(" in formula_upper
            has_a1_ref = "A1" in formula_upper
            has_b5_ref = "B5" in formula_upper
            has_single_quotes = "'" in b8_value
            broken_pattern = '=INDIRECT("Sheet"&A1&".B5")'
            is_broken = (b8_value.replace(" ", "") == broken_pattern.replace(" ", ""))

            if has_indirect and has_a1_ref and has_b5_ref and has_single_quotes and not is_broken:
                print(f"PASS: Component 2 -- Formula is fixed AND dynamic (INDIRECT+A1+B5+quotes) (0.3 pts)")
                total_score += 0.3
            else:
                reasons = []
                if not has_single_quotes or is_broken:
                    reasons.append("missing single-quote fix")
                if not has_indirect:
                    reasons.append("missing INDIRECT")
                if not has_a1_ref:
                    reasons.append("missing A1 ref")
                if not has_b5_ref:
                    reasons.append("missing B5 target")
                print(f"FAIL: Component 2 -- {', '.join(reasons)}. Formula: {b8_value}")
        else:
            print(f"FAIL: Component 2 -- B8 is not a formula. Value: {repr(b8_value)}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Formula resolves to correct value (0.2 points)
    # 'Sheet 1'.B5 = 67800 (West Coast Q1 Revenue)
    # We check the cached/computed value using data_only=True
    try:
        wb_data = openpyxl.load_workbook(file_path, data_only=True)
        ws_data = wb_data['Dashboard']
        b8_computed = ws_data['B8'].value
        print(f"DEBUG: B8 computed (data_only) = {repr(b8_computed)}")

        if b8_computed is not None:
            try:
                numeric_val = float(b8_computed)
                # Expected value: 67800 (from 'Sheet 1'.B5 which is West Coast Q1 Revenue)
                if abs(numeric_val - 67800) < 0.01:
                    print(f"PASS: Component 3 -- B8 resolves to 67800 (correct value) (0.2 pts)")
                    total_score += 0.2
                else:
                    print(f"FAIL: Component 3 -- B8 resolves to {numeric_val}, expected 67800")
            except (ValueError, TypeError):
                print(f"FAIL: Component 3 -- B8 computed value is not numeric: {repr(b8_computed)}")
        else:
            # data_only=True may return None if the file hasn't been opened in LibreOffice
            # In that case, if Component 1 passes (formula has quotes), we give benefit of doubt
            # BUT only if the formula structure is clearly correct
            print(f"INFO: Component 3 -- B8 data_only value is None (file may not have been recalculated)")
            # Fallback: check formula has the fix (single quotes + INDIRECT + A1)
            if b8_value is not None and isinstance(b8_value, str):
                normalized = b8_value.replace(" ", "").upper()
                broken_pat = '=INDIRECT("Sheet"&A1&".B5")'.replace(" ", "").upper()
                is_fixed = ("INDIRECT" in normalized and "'" in b8_value
                            and "A1" in normalized and normalized != broken_pat)
                if is_fixed:
                    print(f"PASS: Component 3 -- Formula structure correct (cached value unavailable) (0.2 pts)")
                    total_score += 0.2
                else:
                    print(f"FAIL: Component 3 -- Cannot verify computed value and formula structure unclear")
            else:
                print(f"FAIL: Component 3 -- No formula and no computed value")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
