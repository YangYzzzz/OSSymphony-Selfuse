"""
Reward Script: Insert a new 'Personnel Records' sheet and protect it with a password
Task ID: calc_gsi_075
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.5): 'Personnel Records' sheet exists in the workbook
  - Component 2 (0.3): 'Personnel Records' sheet has sheet protection enabled
  - Component 3 (0.2): 'Personnel Records' sheet is empty (no data entered before protection)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_075'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: 'Personnel Records' sheet exists (0.5 points)
    # This is the core task requirement — a new sheet must be inserted.
    # On initial_env this sheet does NOT exist, so this correctly scores 0.
    try:
        if 'Personnel Records' in wb.sheetnames:
            print(f"PASS: Component 1 — 'Personnel Records' sheet found in workbook (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — 'Personnel Records' sheet not found. Sheets: {wb.sheetnames}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: 'Personnel Records' sheet has protection enabled (0.3 points)
    # The task requires applying sheet protection with a password.
    # On initial_env the sheet doesn't exist, so this also correctly scores 0.
    try:
        if 'Personnel Records' in wb.sheetnames:
            ws = wb['Personnel Records']
            if ws.protection.sheet:
                # Additionally verify a password hash is set (not just empty protection)
                if ws.protection.password is not None and ws.protection.password != '':
                    print(f"PASS: Component 2 — Sheet protection enabled with password hash '{ws.protection.password}' (0.3 pts)")
                    total_score += 0.3
                else:
                    print(f"FAIL: Component 2 — Sheet protection enabled but no password set (password: {ws.protection.password})")
            else:
                print(f"FAIL: Component 2 — Sheet protection not enabled (protection.sheet={ws.protection.sheet})")
        else:
            print(f"FAIL: Component 2 — Cannot check protection; 'Personnel Records' sheet does not exist")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: 'Personnel Records' sheet is empty / has no data (0.2 points)
    # The task says to protect BEFORE entering any data.
    # On initial_env the sheet doesn't exist, so this also correctly scores 0.
    try:
        if 'Personnel Records' in wb.sheetnames:
            ws = wb['Personnel Records']
            has_data = False
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
                for val in row:
                    if val is not None:
                        has_data = True
                        break
                if has_data:
                    break

            if not has_data:
                print(f"PASS: Component 3 — 'Personnel Records' sheet is empty, protection applied before data entry (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 3 — 'Personnel Records' sheet contains data (should be empty per task requirement)")
        else:
            print(f"FAIL: Component 3 — Cannot check emptiness; 'Personnel Records' sheet does not exist")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'

persist_app_state("libreoffice_calc")

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
