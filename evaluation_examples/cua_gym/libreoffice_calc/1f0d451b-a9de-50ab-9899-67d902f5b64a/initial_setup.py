"""
Initial Setup: Insert a new sheet and protect it with a password
Task ID: calc_gsi_075
Domain: libreoffice_calc

Creates a workbook with existing employee data sheets. Does NOT include
a 'Personnel Records' sheet (that is the task for the agent to create).
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_075'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Employee Directory ---
    ws1 = wb.active
    ws1.title = 'Employee Directory'

    headers = ['Employee ID', 'Full Name', 'Department', 'Position', 'Annual Salary', 'Start Date', 'Email']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    data = [
        ['EMP-001', 'Sarah Chen', 'Engineering', 'Senior Developer', 95000, '2021-03-15', 'sarah.chen@acmecorp.com'],
        ['EMP-002', 'Marcus Johnson', 'Marketing', 'Marketing Manager', 82000, '2020-06-01', 'marcus.j@acmecorp.com'],
        ['EMP-003', 'Priya Patel', 'Finance', 'Financial Analyst', 78000, '2022-01-10', 'priya.p@acmecorp.com'],
        ['EMP-004', 'James O\'Brien', 'Engineering', 'DevOps Engineer', 91000, '2021-09-20', 'james.ob@acmecorp.com'],
        ['EMP-005', 'Linda Martinez', 'Human Resources', 'HR Director', 105000, '2019-04-12', 'linda.m@acmecorp.com'],
        ['EMP-006', 'David Kim', 'Engineering', 'Junior Developer', 68000, '2023-07-01', 'david.k@acmecorp.com'],
        ['EMP-007', 'Rachel Thompson', 'Sales', 'Account Executive', 76000, '2022-11-15', 'rachel.t@acmecorp.com'],
        ['EMP-008', 'Ahmed Hassan', 'Finance', 'Senior Accountant', 85000, '2020-02-28', 'ahmed.h@acmecorp.com'],
        ['EMP-009', 'Emily Nakamura', 'Marketing', 'Content Strategist', 71000, '2023-03-05', 'emily.n@acmecorp.com'],
        ['EMP-010', 'Carlos Rivera', 'Sales', 'Sales Director', 112000, '2018-08-19', 'carlos.r@acmecorp.com'],
        ['EMP-011', 'Sophia Williams', 'Engineering', 'QA Lead', 88000, '2021-05-22', 'sophia.w@acmecorp.com'],
        ['EMP-012', 'Thomas Anderson', 'Human Resources', 'Recruiter', 64000, '2023-09-11', 'thomas.a@acmecorp.com'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Format salary column as currency
    for r in range(2, len(data) + 2):
        ws1.cell(row=r, column=5).number_format = '$#,##0'

    # Adjust column widths
    ws1.column_dimensions['A'].width = 14
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 22
    ws1.column_dimensions['E'].width = 16
    ws1.column_dimensions['F'].width = 14
    ws1.column_dimensions['G'].width = 28

    # Freeze header row
    ws1.freeze_panes = 'A2'

    # --- Sheet 2: Department Summary ---
    ws2 = wb.create_sheet('Department Summary')

    dept_headers = ['Department', 'Headcount', 'Avg Salary', 'Budget Allocation']
    for col, h in enumerate(dept_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    dept_data = [
        ['Engineering', 4, 85500, 450000],
        ['Marketing', 2, 76500, 200000],
        ['Finance', 2, 81500, 180000],
        ['Human Resources', 2, 84500, 175000],
        ['Sales', 2, 94000, 250000],
    ]

    for r, row_data in enumerate(dept_data, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    for r in range(2, len(dept_data) + 2):
        ws2.cell(row=r, column=3).number_format = '$#,##0'
        ws2.cell(row=r, column=4).number_format = '$#,##0'

    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 14
    ws2.column_dimensions['D'].width = 20

    # --- Sheet 3: Quarterly Targets ---
    ws3 = wb.create_sheet('Quarterly Targets')

    qt_headers = ['Quarter', 'Revenue Target', 'Hiring Goal', 'Training Budget']
    for col, h in enumerate(qt_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    qt_data = [
        ['Q1 2025', 1250000, 3, 15000],
        ['Q2 2025', 1400000, 5, 20000],
        ['Q3 2025', 1350000, 2, 12000],
        ['Q4 2025', 1600000, 4, 18000],
    ]

    for r, row_data in enumerate(qt_data, 2):
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r, column=c, value=val)

    for r in range(2, len(qt_data) + 2):
        ws3.cell(row=r, column=2).number_format = '$#,##0'
        ws3.cell(row=r, column=4).number_format = '$#,##0'

    ws3.column_dimensions['A'].width = 14
    ws3.column_dimensions['B'].width = 18
    ws3.column_dimensions['C'].width = 14
    ws3.column_dimensions['D'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
