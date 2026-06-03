"""
Initial Setup: Support ticket tracker for pivot table task
Task ID: calc_adv_pivot_count_004
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date, timedelta
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_adv_pivot_count_004'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Tickets'

    # --- Headers ---
    headers = ['Ticket ID', 'Created Date', 'Priority', 'Team', 'Status', 'Resolution Time']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFFFF')
        cell.alignment = Alignment(horizontal='center')

    # --- Data design: fixed counts per Priority x Team ---
    # Priority x Team counts (totals: each priority=75, each team=75)
    counts = {
        'Critical': {'Frontend': 20, 'Backend': 18, 'DevOps': 19, 'Security': 18},
        'High':     {'Frontend': 19, 'Backend': 20, 'DevOps': 18, 'Security': 18},
        'Medium':   {'Frontend': 18, 'Backend': 19, 'DevOps': 19, 'Security': 19},
        'Low':      {'Frontend': 18, 'Backend': 18, 'DevOps': 19, 'Security': 20},
    }

    # Generate all ticket combinations
    tickets = []
    ticket_num = 1001
    for priority, team_counts in counts.items():
        for team, count in team_counts.items():
            for _ in range(count):
                tickets.append((priority, team))
                ticket_num += 1

    # Shuffle for realistic ordering
    random.shuffle(tickets)

    # Define status pool
    statuses = ['Open', 'In Progress', 'Resolved', 'Closed', 'Pending']
    status_weights = [0.15, 0.25, 0.35, 0.20, 0.05]

    # Base date for created dates
    base_date = date(2024, 1, 2)

    for i, (priority, team) in enumerate(tickets):
        row = i + 2
        ticket_id = f'TKT-{1001 + i}'
        created_date = base_date + timedelta(days=random.randint(0, 364))
        status = random.choices(statuses, weights=status_weights)[0]
        if status in ('Resolved', 'Closed'):
            resolution_time = round(random.uniform(0.5, 72.0), 1)
        else:
            resolution_time = None

        ws.cell(row=row, column=1, value=ticket_id)
        ws.cell(row=row, column=2, value=created_date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=3, value=priority)
        ws.cell(row=row, column=4, value=team)
        ws.cell(row=row, column=5, value=status)
        ws.cell(row=row, column=6, value=resolution_time)

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 18

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Tickets')
    print(f'  Rows: 300 data rows (rows 2-301)')
    print(f'  Columns: Ticket ID, Created Date, Priority, Team, Status, Resolution Time')
    print(f'  No pivot table present.')

create_initial()
