"""
Reward Script: Build a pivot table that counts support tickets by priority x team
Task ID: calc_adv_pivot_count_004
Domain: libreoffice_calc
Scoring:
  Component 1: Pivot sheet exists in workbook (0.2 pts)
  Component 2: Column headers contain the 4 expected teams (0.2 pts)
  Component 3: Row labels contain the 4 expected priority levels (0.2 pts)
  Component 4: Ticket counts are correct for each Priority x Team cell (0.3 pts)
  Component 5: Grand totals are correct (row totals and column totals) (0.1 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_adv_pivot_count_004'

# Expected ground truth values from the task data
# Priority rows x Team columns
EXPECTED_TEAMS = {'Frontend', 'Backend', 'DevOps', 'Security'}
EXPECTED_PRIORITIES = {'Critical', 'High', 'Medium', 'Low'}

# Expected counts derived from scanning the Tickets sheet
# Format: {priority: {team: count}}
EXPECTED_COUNTS = {
    'Critical': {'Frontend': 20, 'Backend': 18, 'DevOps': 19, 'Security': 18},
    'High':     {'Frontend': 19, 'Backend': 20, 'DevOps': 18, 'Security': 18},
    'Medium':   {'Frontend': 18, 'Backend': 19, 'DevOps': 19, 'Security': 19},
    'Low':      {'Frontend': 18, 'Backend': 18, 'DevOps': 19, 'Security': 20},
}
EXPECTED_PRIORITY_TOTALS = {
    'Critical': 75, 'High': 75, 'Medium': 75, 'Low': 75
}
EXPECTED_TEAM_TOTALS = {
    'Frontend': 75, 'Backend': 75, 'DevOps': 75, 'Security': 75
}
EXPECTED_GRAND_TOTAL = 300


def find_pivot_sheet(wb):
    """
    Search for a pivot-like sheet: either named 'Pivot', 'pivot', or a sheet
    that has a grid structure with priority labels on rows and team names on columns.
    Returns the worksheet or None.
    """
    # First, look for a sheet that is NOT 'Tickets' and has rows/cols that suggest a pivot
    for name in wb.sheetnames:
        if name.lower() in ('tickets',):
            continue
        ws = wb[name]
        # A pivot sheet will be small (5-10 rows, 3-7 cols max) vs the 301-row Tickets sheet
        if ws.max_row <= 20 and ws.max_column <= 10:
            return ws
    return None


def extract_pivot_structure(ws):
    """
    Extract headers (row 1) and row labels (column A) from the pivot sheet.
    Returns (col_headers, row_labels, data_dict) where data_dict is {row_label: {col_header: value}}.
    """
    max_row = ws.max_row
    max_col = ws.max_column

    # Row 1 = column headers (skip A1 which is usually the label for the row axis)
    col_headers = {}
    for col in range(2, max_col + 1):
        val = ws.cell(row=1, column=col).value
        if val is not None:
            col_headers[col] = str(val).strip()

    # Column A = row labels (skip row 1)
    row_labels = {}
    for row in range(2, max_row + 1):
        val = ws.cell(row=1 + row - 2 + 1, column=1).value  # same as row
        if val is not None:
            row_labels[row] = str(val).strip()

    # Data cells
    data_dict = {}
    for row, row_label in row_labels.items():
        data_dict[row_label] = {}
        for col, col_header in col_headers.items():
            cell_val = ws.cell(row=row, column=col).value
            data_dict[row_label][col_header] = cell_val

    return col_headers, row_labels, data_dict


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Pivot sheet exists (0.2 points)
    # The initial file only has 'Tickets'. A new pivot-like sheet must exist.
    try:
        pivot_ws = find_pivot_sheet(wb)
        if pivot_ws is not None:
            print(f"PASS: Component 1 — Pivot sheet found: '{pivot_ws.title}' (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 1 — No pivot sheet found. Sheets: {wb.sheetnames}")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print(f"REWARD: {total_score}")
        return total_score

    # Extract structure for remaining components
    try:
        col_headers, row_labels, data_dict = extract_pivot_structure(pivot_ws)
        print(f"  Pivot column headers: {list(col_headers.values())}")
        print(f"  Pivot row labels: {list(row_labels.values())}")
    except Exception as e:
        print(f"ERROR: Cannot parse pivot structure — {e}")
        print(f"REWARD: {total_score}")
        return total_score

    # Component 2: Column headers contain the 4 expected teams (0.2 points)
    # Exclude 'Grand Total' or similar summary column from the check
    try:
        actual_team_headers = {v for v in col_headers.values() if v not in ('Grand Total', 'Total', 'Grand total')}
        found_teams = EXPECTED_TEAMS & actual_team_headers
        if found_teams == EXPECTED_TEAMS:
            print(f"PASS: Component 2 — All 4 team headers found: {sorted(found_teams)} (0.2 pts)")
            total_score += 0.2
        else:
            missing = EXPECTED_TEAMS - actual_team_headers
            extra = actual_team_headers - EXPECTED_TEAMS
            print(f"FAIL: Component 2 — Team headers mismatch. Missing: {missing}, Extra: {extra}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Row labels contain the 4 expected priority levels (0.2 points)
    # Exclude 'Grand Total' summary row
    try:
        actual_priority_labels = {v for v in row_labels.values() if v not in ('Grand Total', 'Total', 'Grand total')}
        found_priorities = EXPECTED_PRIORITIES & actual_priority_labels
        if found_priorities == EXPECTED_PRIORITIES:
            print(f"PASS: Component 3 — All 4 priority labels found: {sorted(found_priorities)} (0.2 pts)")
            total_score += 0.2
        else:
            missing = EXPECTED_PRIORITIES - actual_priority_labels
            extra = actual_priority_labels - EXPECTED_PRIORITIES
            print(f"FAIL: Component 3 — Priority labels mismatch. Missing: {missing}, Extra: {extra}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Correct ticket counts in each Priority x Team cell (0.3 points)
    # All 16 cells must match the expected COUNT values
    try:
        correct_cells = 0
        total_cells = 0
        for priority, team_counts in EXPECTED_COUNTS.items():
            if priority not in data_dict:
                print(f"  WARN: Priority '{priority}' not found in pivot data")
                total_cells += len(team_counts)
                continue
            for team, expected_count in team_counts.items():
                total_cells += 1
                actual_val = data_dict[priority].get(team)
                if actual_val is None:
                    print(f"  WARN: Cell [{priority}][{team}] not found")
                    continue
                try:
                    actual_count = int(actual_val)
                    if actual_count == expected_count:
                        correct_cells += 1
                    else:
                        print(f"  FAIL cell: [{priority}][{team}] expected={expected_count}, got={actual_count}")
                except (ValueError, TypeError):
                    print(f"  FAIL cell: [{priority}][{team}] non-numeric value: {repr(actual_val)}")

        if correct_cells == total_cells and total_cells == 16:
            print(f"PASS: Component 4 — All 16 Priority x Team counts correct ({correct_cells}/16) (0.3 pts)")
            total_score += 0.3
        elif correct_cells > 0:
            partial = round(0.3 * correct_cells / 16, 4)
            print(f"PARTIAL: Component 4 — {correct_cells}/{total_cells} counts correct, partial credit: {partial}")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — 0/{total_cells} counts correct")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Grand totals are correct (0.1 points)
    # Row totals for each priority should be 75, column totals for each team should be 75
    try:
        grand_total_correct = 0
        grand_total_expected = 0

        # Check 'Grand Total' column (row totals) if present
        grand_total_col = None
        for col_idx, col_name in col_headers.items():
            if col_name in ('Grand Total', 'Total', 'Grand total'):
                grand_total_col = col_idx
                break

        # Check 'Grand Total' row (column totals) if present
        grand_total_row = None
        for row_idx, row_name in row_labels.items():
            if row_name in ('Grand Total', 'Total', 'Grand total'):
                grand_total_row = row_idx
                break

        # Verify row grand totals (each priority total = 75)
        for priority, expected_total in EXPECTED_PRIORITY_TOTALS.items():
            grand_total_expected += 1
            if grand_total_col is not None and priority in data_dict:
                gt_header_name = col_headers.get(grand_total_col, '')
                actual_total = data_dict[priority].get(gt_header_name)
                if actual_total is not None:
                    try:
                        if int(actual_total) == expected_total:
                            grand_total_correct += 1
                        else:
                            print(f"  FAIL grand total: row '{priority}' expected {expected_total}, got {actual_total}")
                    except (ValueError, TypeError):
                        print(f"  FAIL grand total: row '{priority}' non-numeric: {repr(actual_total)}")
                else:
                    # No grand total column — check if row sums are correct instead
                    row_sum = 0
                    for team in EXPECTED_TEAMS:
                        v = data_dict[priority].get(team)
                        if v is not None:
                            try:
                                row_sum += int(v)
                            except (ValueError, TypeError):
                                pass
                    if row_sum == expected_total:
                        grand_total_correct += 1
                    else:
                        print(f"  FAIL grand total: row '{priority}' sum={row_sum}, expected {expected_total}")
            else:
                # Compute sum from data_dict
                if priority in data_dict:
                    row_sum = 0
                    for team in EXPECTED_TEAMS:
                        v = data_dict[priority].get(team)
                        if v is not None:
                            try:
                                row_sum += int(v)
                            except (ValueError, TypeError):
                                pass
                    if row_sum == expected_total:
                        grand_total_correct += 1

        # Verify column grand totals (each team total = 75)
        for team, expected_total in EXPECTED_TEAM_TOTALS.items():
            grand_total_expected += 1
            if grand_total_row is not None:
                gt_row_name = row_labels.get(grand_total_row, '')
                actual_total = data_dict.get(gt_row_name, {}).get(team)
                if actual_total is not None:
                    try:
                        if int(actual_total) == expected_total:
                            grand_total_correct += 1
                        else:
                            print(f"  FAIL grand total: col '{team}' expected {expected_total}, got {actual_total}")
                    except (ValueError, TypeError):
                        print(f"  FAIL grand total: col '{team}' non-numeric: {repr(actual_total)}")
                else:
                    # Compute sum from column
                    col_sum = 0
                    for priority in EXPECTED_PRIORITIES:
                        v = data_dict.get(priority, {}).get(team)
                        if v is not None:
                            try:
                                col_sum += int(v)
                            except (ValueError, TypeError):
                                pass
                    if col_sum == expected_total:
                        grand_total_correct += 1
                    else:
                        print(f"  FAIL grand total: col '{team}' sum={col_sum}, expected {expected_total}")
            else:
                # No Grand Total row — compute column sums from data
                col_sum = 0
                for priority in EXPECTED_PRIORITIES:
                    v = data_dict.get(priority, {}).get(team)
                    if v is not None:
                        try:
                            col_sum += int(v)
                        except (ValueError, TypeError):
                            pass
                if col_sum == expected_total:
                    grand_total_correct += 1
                else:
                    print(f"  FAIL grand total: col '{team}' sum={col_sum}, expected {expected_total}")

        ratio = grand_total_correct / grand_total_expected if grand_total_expected > 0 else 0
        if ratio >= 1.0:
            print(f"PASS: Component 5 — All grand totals correct ({grand_total_correct}/{grand_total_expected}) (0.1 pts)")
            total_score += 0.1
        elif ratio >= 0.5:
            partial = round(0.1 * ratio, 4)
            print(f"PARTIAL: Component 5 — {grand_total_correct}/{grand_total_expected} grand totals correct, partial: {partial}")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — Only {grand_total_correct}/{grand_total_expected} grand totals correct")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
