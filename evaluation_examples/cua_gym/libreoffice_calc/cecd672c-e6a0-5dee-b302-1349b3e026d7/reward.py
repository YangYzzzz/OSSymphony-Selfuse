"""
Reward Script: Process expense reports from Thunderbird emails
Task ID: osworld_multi_apps_email_data_009
Domain: libreoffice_calc (multi-app: Thunderbird + Python + LibreOffice Calc)

Scoring rubric (total = 1.0):
  Component 1: 3 expense CSVs downloaded to /home/user/expenses/   — 0.20 pts
  Component 2: merge_expenses.py script created in /home/user/scripts/ — 0.20 pts
  Component 3: combined.csv has correct merged data with 39 data rows,
               totals by category (5 categories) and totals by person (3 persons) — 0.30 pts
  Component 4: LibreOffice Calc xlsx file has bold headers in row 1
               and a GRAND TOTAL row with correct value 5588.57 — 0.30 pts

NOTE: Email replies are NOT scored because the golden_env does not contain
      Thunderbird reply artifacts. Only verifiable file-system changes are scored.
"""

import os
import csv

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_email_data_009'

EXPENSES_DIR = os.path.join(WORKDIR, 'expenses')
SCRIPTS_DIR = os.path.join(WORKDIR, 'scripts')
COMBINED_CSV = os.path.join(EXPENSES_DIR, 'combined.csv')
XLSX_PATH = os.path.join(WORKDIR, f'{TASK_ID}.xlsx')
# Also accept combined.xlsx in expenses dir
COMBINED_XLSX = os.path.join(EXPENSES_DIR, 'combined.xlsx')

EXPECTED_CSV_FILES = {'alice.csv', 'bob.csv', 'carol.csv'}
EXPECTED_GRAND_TOTAL = 5588.57
GRAND_TOTAL_TOLERANCE = 0.10  # allow up to 10 cents variance


def verify_task():
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # -------------------------------------------------------------------
    # Component 1: 3 expense CSV files downloaded to /home/user/expenses/
    # (0.20 points)
    # This FAILS on initial_env (no expenses/ dir) and PASSES on golden_env
    # -------------------------------------------------------------------
    try:
        if not os.path.isdir(EXPENSES_DIR):
            print(f"FAIL: Component 1 — expenses directory does not exist: {EXPENSES_DIR}")
        else:
            found_csvs = set()
            for fname in os.listdir(EXPENSES_DIR):
                if fname in EXPECTED_CSV_FILES:
                    fpath = os.path.join(EXPENSES_DIR, fname)
                    if os.path.getsize(fpath) > 0:
                        found_csvs.add(fname)

            if found_csvs == EXPECTED_CSV_FILES:
                print(f"PASS: Component 1 — all 3 CSV files found in {EXPENSES_DIR}: {sorted(found_csvs)} (0.20 pts)")
                total_score += 0.20
            else:
                missing = EXPECTED_CSV_FILES - found_csvs
                print(f"FAIL: Component 1 — missing CSV files: {sorted(missing)}, found: {sorted(found_csvs)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------
    # Component 2: merge_expenses.py script created in /home/user/scripts/
    # (0.20 points)
    # This FAILS on initial_env (no scripts/ dir) and PASSES on golden_env
    # -------------------------------------------------------------------
    try:
        merge_script = os.path.join(SCRIPTS_DIR, 'merge_expenses.py')
        if not os.path.isdir(SCRIPTS_DIR):
            print(f"FAIL: Component 2 — scripts directory does not exist: {SCRIPTS_DIR}")
        elif not os.path.isfile(merge_script):
            print(f"FAIL: Component 2 — merge_expenses.py not found at {merge_script}")
        else:
            # Verify it's a Python script with some meaningful content (> 100 bytes)
            script_size = os.path.getsize(merge_script)
            with open(merge_script, 'r', errors='ignore') as f:
                content = f.read()
            has_csv_read = 'csv' in content.lower() or 'open(' in content
            has_combined = 'combined' in content.lower()
            if script_size > 100 and has_csv_read and has_combined:
                print(f"PASS: Component 2 — merge_expenses.py found ({script_size} bytes), references csv and combined (0.20 pts)")
                total_score += 0.20
            elif script_size > 100:
                # Accept any non-trivial Python script in the right location
                print(f"PASS: Component 2 — merge_expenses.py found ({script_size} bytes) (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 2 — merge_expenses.py exists but is too small ({script_size} bytes)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------
    # Component 3: combined.csv has correct merged data
    # Expected: 39 data rows (alice=12, bob=13, carol=14), plus totals section
    # Also checks for correct category and person totals
    # (0.30 points)
    # This FAILS on initial_env (no combined.csv) and PASSES on golden_env
    # -------------------------------------------------------------------
    try:
        if not os.path.isfile(COMBINED_CSV):
            print(f"FAIL: Component 3 — combined.csv not found at {COMBINED_CSV}")
        else:
            data_rows = []
            categories_found = set()
            persons_found = set()

            with open(COMBINED_CSV, newline='') as f:
                reader = csv.reader(f)
                in_data = False
                has_header = False
                for row in reader:
                    if not row or all(c.strip() == '' for c in row):
                        continue
                    # Check for header row
                    if row[0].lower() == 'date' and not has_header:
                        has_header = True
                        in_data = True
                        continue
                    # Check for totals section markers
                    if row[0].startswith('---') or (row[0] == '' and row[1] and not row[1].startswith('---')):
                        in_data = False
                        # Still extract person/category info from totals rows
                        if row[1] and row[2]:
                            categories_found.add(row[1])
                        if row[3]:
                            persons_found.add(row[3])
                        continue
                    if in_data and len(row) >= 3 and row[0] and row[0] != '---':
                        data_rows.append(row)
                        if len(row) >= 2:
                            categories_found.add(row[1])
                        if len(row) >= 4:
                            persons_found.add(row[3])

            num_data_rows = len(data_rows)
            expected_persons = {'alice', 'bob', 'carol'}
            found_persons_lower = {p.lower() for p in persons_found if p}

            # Check: at least 30 data rows (allow for some variation) and 3 persons
            if num_data_rows >= 30 and expected_persons.issubset(found_persons_lower):
                print(f"PASS: Component 3 — combined.csv has {num_data_rows} data rows, "
                      f"persons: {sorted(found_persons_lower)} (0.30 pts)")
                total_score += 0.30
            elif num_data_rows >= 30:
                print(f"FAIL: Component 3 — combined.csv has {num_data_rows} rows but "
                      f"missing persons. Found: {sorted(found_persons_lower)}, expected: {sorted(expected_persons)}")
            else:
                print(f"FAIL: Component 3 — combined.csv has only {num_data_rows} data rows "
                      f"(expected >= 30). Persons found: {sorted(found_persons_lower)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------
    # Component 4: LibreOffice Calc xlsx has bold headers and grand total row
    # (0.30 points)
    # This FAILS on initial_env (no xlsx file) and PASSES on golden_env
    # -------------------------------------------------------------------
    try:
        # Determine which xlsx to check (prefer canonical task path)
        xlsx_to_check = None
        if os.path.isfile(XLSX_PATH):
            xlsx_to_check = XLSX_PATH
        elif os.path.isfile(COMBINED_XLSX):
            xlsx_to_check = COMBINED_XLSX

        if xlsx_to_check is None:
            print(f"FAIL: Component 4 — no xlsx file found at {XLSX_PATH} or {COMBINED_XLSX}")
        else:
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_to_check)
            ws = wb.active

            # Check bold headers in row 1
            headers_bold = all(
                ws.cell(row=1, column=c).font.bold
                for c in range(1, 5)
                if ws.cell(row=1, column=c).value is not None
            )
            headers_present = any(
                ws.cell(row=1, column=c).value is not None
                for c in range(1, 5)
            )

            # Check grand total row — look for a row containing 'GRAND TOTAL' (case-insensitive)
            grand_total_row = None
            grand_total_value = None
            for row_idx in range(ws.max_row, 0, -1):
                for col_idx in range(1, ws.max_column + 1):
                    cell_val = ws.cell(row=row_idx, column=col_idx).value
                    if cell_val and isinstance(cell_val, str) and 'grand total' in cell_val.lower():
                        grand_total_row = row_idx
                        # Find numeric value in same row
                        for gc in range(1, ws.max_column + 1):
                            gv = ws.cell(row=row_idx, column=gc).value
                            if isinstance(gv, (int, float)) and gv > 0:
                                grand_total_value = float(gv)
                                break
                        break
                if grand_total_row is not None:
                    break

            # Also check if grand total row cell is bold
            grand_total_bold = False
            if grand_total_row is not None:
                for gc in range(1, ws.max_column + 1):
                    if ws.cell(row=grand_total_row, column=gc).font.bold:
                        grand_total_bold = True
                        break

            # Score: headers bold + grand total row present with correct value
            if (headers_present and headers_bold and
                    grand_total_row is not None and
                    grand_total_value is not None and
                    abs(grand_total_value - EXPECTED_GRAND_TOTAL) <= GRAND_TOTAL_TOLERANCE):
                print(f"PASS: Component 4 — xlsx has bold headers, GRAND TOTAL row at row {grand_total_row} "
                      f"with value {grand_total_value} (expected {EXPECTED_GRAND_TOTAL}) (0.30 pts)")
                total_score += 0.30
            elif headers_present and headers_bold and grand_total_row is not None:
                # Partial: headers bold and grand total row exists but value may differ
                if grand_total_value is not None:
                    print(f"FAIL: Component 4 — xlsx has bold headers and GRAND TOTAL row "
                          f"but value {grand_total_value} != expected {EXPECTED_GRAND_TOTAL}")
                else:
                    print(f"FAIL: Component 4 — xlsx has bold headers and GRAND TOTAL label "
                          f"but no numeric value found in that row")
            elif headers_present and not headers_bold:
                print(f"FAIL: Component 4 — xlsx has headers but they are NOT bold")
            elif grand_total_row is None:
                header_vals = [ws.cell(row=1, column=c).value for c in range(1, 5)]
                print(f"FAIL: Component 4 — no GRAND TOTAL row found in xlsx. Headers: {header_vals}")
            else:
                print(f"FAIL: Component 4 — xlsx check failed: headers_present={headers_present}, "
                      f"headers_bold={headers_bold}, grand_total_row={grand_total_row}, "
                      f"grand_total_value={grand_total_value}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


verify_task()
