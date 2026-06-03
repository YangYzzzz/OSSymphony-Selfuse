"""
Initial Setup: Apply Data Bar conditional formatting to score column
Task ID: calc_gg1_029
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_029'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leaderboard"

    # --- Headers ---
    headers = ["Player Name", "Level", "Wins", "Total Score"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Player data (40 rows: D2:D41) ---
    player_names = [
        "DragonSlayer99", "PixelQueen", "NightHawk_X", "SilverBullet",
        "CrypticWolf", "BlazeFury", "IronClad_7", "StormRider",
        "VenomStrike", "PhantomAce", "LunarEclipse", "TurboFox",
        "ShadowMercer", "CosmicDrift", "ThunderBolt3", "NeonViper",
        "FrostByte", "GalacticNova", "MysticRaven", "CyberPunk42",
        "SteelTalon", "ArcticWind", "BlazeKnight", "QuantumLeap",
        "EmberFang", "NovaStrike", "ZeroGravity", "WildCard_X",
        "DarkMatter", "SonicBoom", "TitanForge", "VortexRush",
        "CrimsonTide", "AshPhoenix", "OmegaWolf", "EchoStorm",
        "PrismShade", "RogueAgent", "InfernoKing", "SpectrumX",
    ]

    # Predefined scores spanning ~500 to ~98,000
    scores = [
        97500, 91200, 88750, 85300, 82100,
        78600, 75400, 71800, 68200, 64500,
        61300, 58900, 55100, 52400, 49700,
        46800, 43200, 40100, 37500, 34800,
        31600, 28900, 26300, 23700, 21500,
        18900, 16400, 14200, 11800, 9500,
        8200, 7100, 5800, 4600, 3500,
        2800, 2100, 1400, 850, 520,
    ]

    random.seed(42)
    levels = [random.randint(1, 100) for _ in range(40)]
    wins_list = [random.randint(5, 500) for _ in range(40)]

    data_font = Font(name="Calibri", size=11)
    data_align = Alignment(horizontal="left", vertical="center")
    num_align = Alignment(horizontal="right", vertical="center")

    for i in range(40):
        row = i + 2
        # Player Name
        c = ws.cell(row=row, column=1, value=player_names[i])
        c.font = data_font
        c.alignment = data_align
        c.border = thin_border

        # Level
        c = ws.cell(row=row, column=2, value=levels[i])
        c.font = data_font
        c.alignment = num_align
        c.border = thin_border

        # Wins
        c = ws.cell(row=row, column=3, value=wins_list[i])
        c.font = data_font
        c.alignment = num_align
        c.border = thin_border

        # Total Score
        c = ws.cell(row=row, column=4, value=scores[i])
        c.font = data_font
        c.alignment = num_align
        c.border = thin_border
        c.number_format = '#,##0'

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 16

    # Freeze header row
    ws.freeze_panes = "A2"

    # NO conditional formatting in initial state
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
