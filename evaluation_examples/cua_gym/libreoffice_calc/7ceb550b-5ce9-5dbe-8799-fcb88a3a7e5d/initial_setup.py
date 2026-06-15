"""
Initial Setup: Create dates.xlsx and events.ods on the Desktop with matching row counts.
Task ID: osworld_multi_apps_terminal_calc_006
Domain: libreoffice_calc (multi-app: terminal + calc)
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user/Desktop'
TASK_ID = 'osworld_multi_apps_terminal_calc_006'
DATES_FILE = f'{WORKDIR}/dates.xlsx'
EVENTS_FILE = f'{WORKDIR}/events.ods'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_dates_xlsx():
    """Create dates.xlsx with a single 'Date' column containing realistic dates."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dates"

    # Header
    ws.cell(row=1, column=1, value="Date")

    # Realistic dates (unsorted intentionally so the task requires sorting)
    dates = [
        "2025-04-15",
        "2025-01-08",
        "2025-06-22",
        "2025-03-03",
        "2025-07-10",
        "2025-02-18",
        "2025-05-30",
        "2025-08-05",
        "2025-09-12",
        "2025-11-27",
        "2025-10-04",
        "2025-12-19",
    ]

    for r, date_val in enumerate(dates, 2):
        ws.cell(row=r, column=1, value=date_val)

    wb.save(DATES_FILE)
    print(f'Created: {DATES_FILE}')


def create_events_ods():
    """Create events.ods with a single 'Event' column using realistic event names.

    We use openpyxl to create an .xlsx first then convert to .ods via LibreOffice headless,
    because openpyxl cannot write .ods directly.
    """
    # First create a temporary xlsx
    tmp_xlsx = f'/tmp/events_tmp.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Events"

    # Header
    ws.cell(row=1, column=1, value="Event")

    # Realistic event names (one per date row, same order as dates above)
    events = [
        "Spring Product Launch",
        "New Year Kickoff Meeting",
        "Midsummer Tech Conference",
        "Q1 Strategy Review",
        "Annual Team Retreat",
        "Quarterly Budget Planning",
        "Customer Appreciation Day",
        "Engineering Summit",
        "Fall Innovation Workshop",
        "Thanksgiving Charity Gala",
        "October Sales Seminar",
        "Year-End Awards Ceremony",
    ]

    for r, event_name in enumerate(events, 2):
        ws.cell(row=r, column=1, value=event_name)

    wb.save(tmp_xlsx)
    print(f'Temporary xlsx created: {tmp_xlsx}')

    # Convert xlsx to ods using LibreOffice headless
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    result = subprocess.run(
        [
            "libreoffice",
            "--headless",
            "--convert-to", "ods",
            "--outdir", WORKDIR,
            tmp_xlsx,
        ],
        env=env,
        capture_output=True,
        text=True,
        timeout=60,
    )
    print(f'LibreOffice conversion stdout: {result.stdout}')
    print(f'LibreOffice conversion stderr: {result.stderr}')

    # The converted file will be named events_tmp.ods; rename it to events.ods
    converted_path = f'{WORKDIR}/events_tmp.ods'
    if os.path.exists(converted_path):
        os.rename(converted_path, EVENTS_FILE)
        print(f'Renamed to: {EVENTS_FILE}')
    else:
        # Try alternative path if conversion placed it elsewhere
        alt_path = f'/tmp/events_tmp.ods'
        if os.path.exists(alt_path):
            import shutil
            shutil.copy2(alt_path, EVENTS_FILE)
            print(f'Copied from alt path to: {EVENTS_FILE}')
        else:
            print(f'WARNING: Conversion output not found at {converted_path} or {alt_path}')
            print(f'Listing /tmp: {os.listdir("/tmp")}')
            print(f'Listing {WORKDIR}: {os.listdir(WORKDIR)}')

    # Clean up temp file
    if os.path.exists(tmp_xlsx):
        os.remove(tmp_xlsx)


def create_initial():
    # Ensure Desktop directory exists
    os.makedirs(WORKDIR, exist_ok=True)

    # Remove any leftover schedule.csv from previous runs (idempotent)
    schedule_csv = f'{WORKDIR}/schedule.csv'
    if os.path.exists(schedule_csv):
        os.remove(schedule_csv)
        print(f'Removed existing schedule.csv for clean initial state')

    # Create the two input files
    create_dates_xlsx()
    create_events_ods()

    print(f'Initial files created:')
    print(f'  {DATES_FILE}')
    print(f'  {EVENTS_FILE}')

    # GUI-ready startup: open a terminal (task expects terminal to be open)
    launch_gui('gnome-terminal', delay_sec=2.0)
    print('GUI_READY: launched gnome-terminal with DISPLAY=:0')


create_initial()
