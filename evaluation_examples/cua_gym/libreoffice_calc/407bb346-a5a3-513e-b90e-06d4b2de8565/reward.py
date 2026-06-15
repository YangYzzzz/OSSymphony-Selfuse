"""
Reward Script: Fellowship Grant Pass Rate Table
Task ID: osworld_multi_apps_ecs_multi_report_006
Domain: libreoffice_calc

Task: Read 5 PDF annual reports (2019-2023) from ~/Documents/Fellowships/,
compile a pass rate table in LibreOffice Calc with disciplines as rows and
years as columns, save as 'fellowship_by_discipline.xlsx' on Desktop.

Scoring Rubric:
  Component 1: File exists at ~/Desktop/fellowship_by_discipline.xlsx   (gate)
  Component 2: Correct spreadsheet structure (5 discipline rows, 5 year columns)  0.30
  Component 3: All discipline names present in column A                  0.30
  Component 4: Year headers (2019-2023) present in row 1                 0.10
  Component 5: Pass rate values numerically correct (within tolerance)   0.30
  Total: 1.00
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_ecs_multi_report_006'
FILE_PATH = f'{WORKDIR}/Desktop/fellowship_by_discipline.xlsx'

# Expected pass rate data extracted from PDFs (as decimal, e.g. 0.3790 = 37.90%)
# discipline -> {year: pass_rate}
EXPECTED_DATA = {
    'Biology':          {2019: 0.3790, 2020: 0.3969, 2021: 0.3898, 2022: 0.4048, 2023: 0.4148},
    'Chemistry':        {2019: 0.3673, 2020: 0.3883, 2021: 0.4000, 2022: 0.4112, 2023: 0.4196},
    'Physics':          {2019: 0.3563, 2020: 0.3736, 2021: 0.3929, 2022: 0.4045, 2023: 0.4194},
    'Computer Science': {2019: 0.4018, 2020: 0.4219, 2021: 0.4336, 2022: 0.4459, 2023: 0.4643},
    'Mathematics':      {2019: 0.3421, 2020: 0.3537, 2021: 0.3544, 2022: 0.3765, 2023: 0.3846},
}

EXPECTED_DISCIPLINES = ['Biology', 'Chemistry', 'Physics', 'Computer Science', 'Mathematics']
EXPECTED_YEARS = [2019, 2020, 2021, 2022, 2023]

TOLERANCE = 0.005  # allow 0.5% tolerance for rounding differences


def find_header_row_and_col(ws):
    """
    Find the header row that contains year values and the discipline column.
    Returns (header_row, discipline_col, year_to_col_map) or None if not found.
    """
    for row in ws.iter_rows():
        year_cols = {}
        discipline_col = None
        for cell in row:
            val = cell.value
            if val is None:
                continue
            # Check for year values (as int or string)
            try:
                year_int = int(str(val).strip())
                if 2019 <= year_int <= 2023:
                    year_cols[year_int] = cell.column
            except (ValueError, TypeError):
                pass
            # Check for discipline header (case-insensitive)
            if isinstance(val, str) and val.strip().lower() == 'discipline':
                discipline_col = cell.column
        if len(year_cols) >= 5 and discipline_col is not None:
            return cell.row, discipline_col, year_cols
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Gate: file must exist
    if not os.path.exists(file_path):
        print(f"CRITICAL: File not found: {file_path}")
        print("REWARD: 0.0")
        return 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # ----- Component 1: Correct spreadsheet structure -----
    # Should have at least 6 rows (1 header + 5 disciplines) and 6 columns (1 discipline + 5 years)
    # (0.30 points)
    try:
        data_rows = 0
        data_cols = 0
        for row in ws.iter_rows():
            non_empty = [c for c in row if c.value is not None]
            if len(non_empty) >= 5:
                data_rows += 1
                data_cols = max(data_cols, len(non_empty))

        if data_rows >= 6 and data_cols >= 6:
            print(f"PASS: Component 1 — Spreadsheet has {data_rows} data rows and {data_cols} non-empty columns (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 — Expected at least 6 data rows and 6 columns, found {data_rows} rows and {data_cols} columns")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ----- Component 2: Correct discipline names in column A -----
    # (0.30 points)
    try:
        found_disciplines = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    val = cell.value.strip()
                    if val in EXPECTED_DISCIPLINES:
                        found_disciplines.append(val)
                        break

        found_set = set(found_disciplines)
        expected_set = set(EXPECTED_DISCIPLINES)
        matched = found_set & expected_set
        missing = expected_set - found_set

        if len(matched) == 5:
            print(f"PASS: Component 2 — All 5 discipline names found: {sorted(matched)} (0.30 pts)")
            total_score += 0.30
        elif len(matched) >= 3:
            partial = round(0.30 * len(matched) / 5, 4)
            print(f"PARTIAL: Component 2 — {len(matched)}/5 discipline names found, missing: {missing} ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {len(matched)}/5 discipline names found, missing: {missing}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ----- Component 3: Year headers 2019-2023 in row 1 -----
    # (0.10 points)
    try:
        found_years = []
        for row in ws.iter_rows(max_row=3):  # headers should be in first 3 rows
            for cell in row:
                try:
                    year_int = int(str(cell.value).strip())
                    if year_int in EXPECTED_YEARS:
                        found_years.append(year_int)
                except (ValueError, TypeError):
                    pass
            if len(set(found_years)) == 5:
                break

        found_years_set = set(found_years)
        if len(found_years_set) == 5:
            print(f"PASS: Component 3 — All year headers 2019-2023 found (0.10 pts)")
            total_score += 0.10
        elif len(found_years_set) >= 3:
            partial = round(0.10 * len(found_years_set) / 5, 4)
            print(f"PARTIAL: Component 3 — {len(found_years_set)}/5 year headers found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {len(found_years_set)}/5 year headers found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ----- Component 4: Pass rate values are numerically correct -----
    # (0.30 points)
    try:
        # Find header row and mapping
        header_info = find_header_row_and_col(ws)
        if header_info is None:
            print("FAIL: Component 4 — Could not find header row with year columns and discipline column")
        else:
            header_row, discipline_col, year_to_col = header_info

            # Build mapping from discipline name -> row number
            discipline_rows = {}
            for r in range(header_row + 1, ws.max_row + 1):
                cell_val = ws.cell(row=r, column=discipline_col).value
                if isinstance(cell_val, str) and cell_val.strip() in EXPECTED_DISCIPLINES:
                    discipline_rows[cell_val.strip()] = r

            # Check each discipline-year combination
            correct_count = 0
            total_count = 0
            errors = []
            for disc, year_data in EXPECTED_DATA.items():
                if disc not in discipline_rows:
                    errors.append(f"{disc}: row not found")
                    total_count += 5
                    continue
                row_num = discipline_rows[disc]
                for year, expected_rate in year_data.items():
                    total_count += 1
                    if year not in year_to_col:
                        errors.append(f"{disc}/{year}: column not found")
                        continue
                    col_num = year_to_col[year]
                    actual_val = ws.cell(row=row_num, column=col_num).value
                    if actual_val is None:
                        errors.append(f"{disc}/{year}: value is None")
                        continue
                    try:
                        actual_float = float(actual_val)
                        # Handle percentage vs decimal representation
                        # PDFs show 37.90% → could be stored as 0.3790 or 37.90
                        if actual_float > 1.0:
                            actual_float = actual_float / 100.0
                        if abs(actual_float - expected_rate) <= TOLERANCE:
                            correct_count += 1
                        else:
                            errors.append(f"{disc}/{year}: expected ~{expected_rate:.4f}, got {actual_val!r} (={actual_float:.4f})")
                    except (ValueError, TypeError) as ve:
                        errors.append(f"{disc}/{year}: cannot convert {actual_val!r} to float: {ve}")

            accuracy = correct_count / total_count if total_count > 0 else 0
            pts = round(0.30 * accuracy, 4)
            if correct_count == total_count:
                print(f"PASS: Component 4 — All {correct_count}/{total_count} pass rate values correct (0.30 pts)")
                total_score += 0.30
            elif correct_count >= total_count * 0.7:
                print(f"PARTIAL: Component 4 — {correct_count}/{total_count} pass rates correct ({pts} pts)")
                if errors:
                    print(f"  Errors: {errors[:5]}")
                total_score += pts
            else:
                print(f"FAIL: Component 4 — Only {correct_count}/{total_count} pass rates correct")
                if errors:
                    print(f"  Errors: {errors[:10]}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score:.4f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Run verification
if not os.path.exists(FILE_PATH):
    print(f"File not found: {FILE_PATH}")
    print("REWARD: 0.0")
else:
    verify_task(FILE_PATH)
