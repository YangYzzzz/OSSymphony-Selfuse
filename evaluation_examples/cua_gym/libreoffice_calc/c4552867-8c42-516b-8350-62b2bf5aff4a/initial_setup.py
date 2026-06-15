"""
Initial Setup: HR Working Schedule - Shift schedule template for next month
Task ID: calc_hr_working_schedule_038
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_working_schedule_038'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Schedule ---
    ws = wb.active
    ws.title = 'Schedule'

    # Row 1 headers
    headers = ['Employee', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Employee names (realistic, 25 employees)
    employees = [
        'Sarah Chen',
        'Marcus Johnson',
        'Priya Patel',
        'Daniel Kim',
        'Laura Martinez',
        'James O\'Brien',
        'Aisha Nwosu',
        'Ryan Thompson',
        'Mei Lin',
        'Carlos Ramirez',
        'Emma Fitzgerald',
        'Noah Williams',
        'Fatima Al-Hassan',
        'Liam Nakamura',
        'Isabella Costa',
        'Ethan Okonkwo',
        'Sofia Bergmann',
        'Jack Sullivan',
        'Yuna Park',
        'Tyler Brooks',
        'Amara Diallo',
        'Connor Walsh',
        'Zara Hussain',
        'Benjamin Lefebvre',
        'Natalie Kowalski',
    ]

    # Rows 2-26: employee names in column A, columns B-H empty (no shifts assigned)
    for row_idx, name in enumerate(employees, 2):
        ws.cell(row=row_idx, column=1, value=name)
        # Columns B-H intentionally left empty (no shifts assigned yet)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 22
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col_letter].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Schedule')
    print(f'  Rows: 1 header + 25 employee rows (rows 2-26)')
    print(f'  Columns B-H: empty (no shifts assigned)')
    print(f'  No data validation, no conditional formatting, no bold on names')

create_initial()
