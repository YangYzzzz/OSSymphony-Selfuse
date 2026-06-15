"""
Reward Script: Set up shift schedule template with dropdowns, alternating row colors, and bold employee names
Task ID: calc_hr_working_schedule_038
Domain: libreoffice_calc
Scoring:
  Component 1 (0.40): Data validation dropdown on B2:H26 with Morning/Afternoon/Night/Off options
  Component 2 (0.30): Bold formatting on employee names in A2:A26
  Component 3 (0.30): Alternating row colors via conditional formatting on A2:H26
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_working_schedule_038'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Schedule' sheet must exist
    if 'Schedule' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Schedule' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Schedule']

    # Component 1: Data validation dropdown on B2:H26 (0.40 points)
    # Must have a list validation with Morning/Afternoon/Night/Off covering B2:H26
    try:
        validations = ws.data_validations.dataValidation
        dv_found = False
        dv_correct_options = False
        dv_correct_range = False

        for dv in validations:
            if dv.type == 'list':
                # Check formula1 contains the required options
                formula = dv.formula1 or ''
                formula_clean = formula.strip('"').strip("'")
                required_options = ['Morning', 'Afternoon', 'Night', 'Off']
                options_present = all(opt in formula_clean for opt in required_options)

                # Check the sqref covers B2:H26
                sqref_str = str(dv.sqref)
                covers_range = 'B2:H26' in sqref_str

                if options_present and covers_range:
                    dv_found = True
                    dv_correct_options = True
                    dv_correct_range = True
                    break
                elif options_present:
                    dv_found = True
                    dv_correct_options = True

        if dv_found and dv_correct_options and dv_correct_range:
            print(f"PASS: Component 1 — Data validation dropdown on B2:H26 with Morning/Afternoon/Night/Off (0.40 pts)")
            total_score += 0.40
        elif dv_found and dv_correct_options:
            print(f"FAIL: Component 1 — Dropdown options are correct but sqref does not include B2:H26 (found: {sqref_str})")
        elif len(validations) > 0:
            print(f"FAIL: Component 1 — Data validation exists but wrong type or options. formula1={validations[0].formula1}")
        else:
            print(f"FAIL: Component 1 — No data validation found on the sheet")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Bold formatting on employee names in A2:A26 (0.30 points)
    # All 25 cells in A2:A26 must have bold=True
    try:
        bold_cells = 0
        non_bold_cells = []
        for r in range(2, 27):
            cell = ws.cell(row=r, column=1)
            if cell.font and cell.font.bold:
                bold_cells += 1
            else:
                non_bold_cells.append(f"A{r}")

        if bold_cells == 25:
            print(f"PASS: Component 2 — All 25 employee names in A2:A26 are bold (0.30 pts)")
            total_score += 0.30
        elif bold_cells > 0:
            print(f"FAIL: Component 2 — Only {bold_cells}/25 cells in A2:A26 are bold. Non-bold: {non_bold_cells[:5]}...")
        else:
            print(f"FAIL: Component 2 — No bold cells found in A2:A26")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Alternating row colors via conditional formatting on A2:H26 (0.30 points)
    # Requires two formula-based rules on A2:H26:
    #   - MOD(ROW(),2)=1 → fill #DCE6F1 (odd rows: light blue)
    #   - MOD(ROW(),2)=0 → fill #FFFFFF (even rows: white)
    try:
        cf_rules = ws.conditional_formatting
        found_odd_rule = False
        found_even_rule = False
        odd_color_ok = False
        even_color_ok = False

        for cf in cf_rules:
            cf_range_str = str(cf)
            # Check that the rule applies to A2:H26 region
            applies_to_target = 'A2' in cf_range_str or 'H26' in cf_range_str

            for rule in cf.rules:
                if rule.type == 'expression' and rule.formula:
                    formula_str = rule.formula[0] if isinstance(rule.formula, list) else rule.formula
                    formula_upper = formula_str.upper().replace(' ', '')

                    # Check for MOD(ROW(),2)=1 rule (odd rows → light blue #DCE6F1)
                    is_odd_formula = ('MOD(ROW(),2)=1' in formula_upper)
                    # Check for MOD(ROW(),2)=0 rule (even rows → white #FFFFFF)
                    is_even_formula = ('MOD(ROW(),2)=0' in formula_upper)

                    if is_odd_formula:
                        found_odd_rule = True
                        # Verify the fill color is #DCE6F1 (ARGB: FFDCE6F1)
                        try:
                            fill_color = rule.dxf.fill.fgColor.rgb
                            if fill_color.upper() in ('FFDCE6F1', 'DCE6F1'):
                                odd_color_ok = True
                                print(f"  INFO: Odd-row rule found, color={fill_color}")
                            else:
                                print(f"  WARN: Odd-row rule found but color is {fill_color}, expected FFDCE6F1")
                        except Exception as ce:
                            print(f"  WARN: Could not read odd-row fill color: {ce}")
                            odd_color_ok = True  # rule exists even if color check failed

                    if is_even_formula:
                        found_even_rule = True
                        # Verify the fill color is #FFFFFF (ARGB: FFFFFFFF)
                        try:
                            fill_color = rule.dxf.fill.fgColor.rgb
                            if fill_color.upper() in ('FFFFFFFF', 'FFFFFF'):
                                even_color_ok = True
                                print(f"  INFO: Even-row rule found, color={fill_color}")
                            else:
                                print(f"  WARN: Even-row rule found but color is {fill_color}, expected FFFFFFFF")
                        except Exception as ce:
                            print(f"  WARN: Could not read even-row fill color: {ce}")
                            even_color_ok = True  # rule exists even if color check failed

        if found_odd_rule and found_even_rule and odd_color_ok and even_color_ok:
            print(f"PASS: Component 3 — Alternating row color conditional formatting on A2:H26 with correct colors (0.30 pts)")
            total_score += 0.30
        elif found_odd_rule and found_even_rule:
            print(f"FAIL: Component 3 — Both CF rules found but colors incorrect. odd_ok={odd_color_ok}, even_ok={even_color_ok}")
        elif found_odd_rule or found_even_rule:
            print(f"FAIL: Component 3 — Only one of the two CF rules found. odd={found_odd_rule}, even={found_even_rule}")
        else:
            print(f"FAIL: Component 3 — No alternating row conditional formatting rules found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
