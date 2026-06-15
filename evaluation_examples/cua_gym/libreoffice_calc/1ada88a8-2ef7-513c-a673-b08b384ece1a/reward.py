"""
Reward Script: Create hyperlinks on Table of Contents sheet navigating to other sheets
Task ID: calc_gsi_082
Domain: libreoffice_calc
Scoring:
  Component 1: A5 hyperlink to 'Q1 Sales' sheet (0.25)
  Component 2: A6 hyperlink to 'Q2 Sales' sheet (0.25)
  Component 3: A7 hyperlink to 'Inventory' sheet (0.25)
  Component 4: A8 hyperlink to 'Employees' sheet (0.25)
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_082'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def check_hyperlink_to_sheet(ws, cell_coord, expected_sheet_name):
    """
    Check if cell has a hyperlink targeting the expected sheet.
    Hyperlinks to internal sheets use targets like "#'Sheet Name'.A1"
    or location like "'Sheet Name'.A1".
    Returns True if a hyperlink pointing to the expected sheet is found.
    """
    cell = ws[cell_coord]
    hl = cell.hyperlink
    if hl is None:
        return False

    # Check target field — common pattern: "#'Sheet Name'.A1" or "#Sheet Name.A1"
    target = hl.target or ''
    location = hl.location or ''

    # Normalize: the sheet name could appear in target or location
    # Valid patterns:
    #   target="#'Q1 Sales'.A1"
    #   location="'Q1 Sales'.A1"
    #   target="#Q1 Sales.A1" (no quotes if no spaces, but these have spaces)
    combined = (target + ' ' + location).lower()
    expected_lower = expected_sheet_name.lower()

    # Check if the expected sheet name appears in either target or location
    if expected_lower in combined:
        return True

    return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Table of Contents' sheet must exist
    if 'Table of Contents' not in wb.sheetnames:
        print("FAIL: 'Table of Contents' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Table of Contents']

    # Expected hyperlinks: cell -> target sheet name
    expected_links = {
        'A5': 'Q1 Sales',
        'A6': 'Q2 Sales',
        'A7': 'Inventory',
        'A8': 'Employees',
    }

    # Component 1: A5 hyperlink to 'Q1 Sales' (0.25 points)
    try:
        if check_hyperlink_to_sheet(ws, 'A5', 'Q1 Sales'):
            print(f"PASS: Component 1 — A5 has hyperlink to 'Q1 Sales' (0.25 pts)")
            total_score += 0.25
        else:
            hl = ws['A5'].hyperlink
            hl_info = f"target={repr(hl.target)}, location={repr(hl.location)}" if hl else "no hyperlink"
            print(f"FAIL: Component 1 — A5 expected hyperlink to 'Q1 Sales', found: {hl_info}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: A6 hyperlink to 'Q2 Sales' (0.25 points)
    try:
        if check_hyperlink_to_sheet(ws, 'A6', 'Q2 Sales'):
            print(f"PASS: Component 2 — A6 has hyperlink to 'Q2 Sales' (0.25 pts)")
            total_score += 0.25
        else:
            hl = ws['A6'].hyperlink
            hl_info = f"target={repr(hl.target)}, location={repr(hl.location)}" if hl else "no hyperlink"
            print(f"FAIL: Component 2 — A6 expected hyperlink to 'Q2 Sales', found: {hl_info}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: A7 hyperlink to 'Inventory' (0.25 points)
    try:
        if check_hyperlink_to_sheet(ws, 'A7', 'Inventory'):
            print(f"PASS: Component 3 — A7 has hyperlink to 'Inventory' (0.25 pts)")
            total_score += 0.25
        else:
            hl = ws['A7'].hyperlink
            hl_info = f"target={repr(hl.target)}, location={repr(hl.location)}" if hl else "no hyperlink"
            print(f"FAIL: Component 3 — A7 expected hyperlink to 'Inventory', found: {hl_info}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: A8 hyperlink to 'Employees' (0.25 points)
    try:
        if check_hyperlink_to_sheet(ws, 'A8', 'Employees'):
            print(f"PASS: Component 4 — A8 has hyperlink to 'Employees' (0.25 pts)")
            total_score += 0.25
        else:
            hl = ws['A8'].hyperlink
            hl_info = f"target={repr(hl.target)}, location={repr(hl.location)}" if hl else "no hyperlink"
            print(f"FAIL: Component 4 — A8 expected hyperlink to 'Employees', found: {hl_info}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
