"""
Initial Setup: Create a workbook with multiple sheets and a Table of Contents sheet
(without hyperlinks - the agent's task is to add them).
Task ID: calc_gsi_082
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_082'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Common styles ---
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # =========================================================
    # Sheet 1: Table of Contents (plain text, NO hyperlinks)
    # =========================================================
    ws_toc = wb.active
    ws_toc.title = "Table of Contents"

    # Title row
    ws_toc.merge_cells("A1:B1")
    ws_toc["A1"] = "Workbook Navigation"
    ws_toc["A1"].font = Font(name="Arial", size=16, bold=True, color="2F5496")
    ws_toc["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Subtitle
    ws_toc.merge_cells("A2:B2")
    ws_toc["A2"] = "Click links below to navigate to each section"
    ws_toc["A2"].font = Font(name="Arial", size=10, italic=True, color="666666")
    ws_toc["A2"].alignment = Alignment(horizontal="center")

    # Headers for TOC
    toc_headers = ["Sheet Name", "Description"]
    for col, h in enumerate(toc_headers, 1):
        cell = ws_toc.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # TOC entries (plain text only - agent must add hyperlinks)
    toc_data = [
        ["Q1 Sales", "First quarter sales figures (Jan-Mar 2025)"],
        ["Q2 Sales", "Second quarter sales figures (Apr-Jun 2025)"],
        ["Inventory", "Current product inventory and stock levels"],
        ["Employees", "Employee directory and department info"],
    ]
    for r, (sheet_name, desc) in enumerate(toc_data, 5):
        cell_a = ws_toc.cell(row=r, column=1, value=sheet_name)
        cell_a.font = Font(name="Arial", size=11)
        cell_a.border = thin_border

        cell_b = ws_toc.cell(row=r, column=2, value=desc)
        cell_b.font = Font(name="Arial", size=11)
        cell_b.border = thin_border

    ws_toc.column_dimensions["A"].width = 22
    ws_toc.column_dimensions["B"].width = 50

    # =========================================================
    # Sheet 2: Q1 Sales
    # =========================================================
    ws_q1 = wb.create_sheet("Q1 Sales")
    q1_headers = ["Sales Rep", "Region", "January", "February", "March", "Q1 Total"]
    for col, h in enumerate(q1_headers, 1):
        cell = ws_q1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    q1_data = [
        ["Sarah Chen", "Northeast", 45230, 38920, 52100, 136250],
        ["Marcus Johnson", "Southeast", 38750, 42300, 35680, 116730],
        ["Elena Rodriguez", "West Coast", 52100, 48750, 55320, 156170],
        ["David Kim", "Midwest", 31200, 35400, 29800, 96400],
        ["Priya Patel", "Southwest", 44800, 39200, 47600, 131600],
        ["James O'Brien", "Northeast", 37600, 41200, 38900, 117700],
        ["Aisha Williams", "Southeast", 29500, 33100, 36200, 98800],
        ["Carlos Mendez", "West Coast", 48300, 51200, 46800, 146300],
        ["Lisa Chang", "Midwest", 35700, 32400, 38100, 106200],
        ["Robert Taylor", "Southwest", 42100, 44500, 40300, 126900],
        ["Nina Kowalski", "Northeast", 39800, 37600, 43200, 120600],
        ["Ahmed Hassan", "Southeast", 33400, 36800, 31500, 101700],
    ]
    for r, row_data in enumerate(q1_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_q1.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c >= 3:
                cell.number_format = '$#,##0'

    for col_letter in ["A", "B"]:
        ws_q1.column_dimensions[col_letter].width = 18
    for col_letter in ["C", "D", "E", "F"]:
        ws_q1.column_dimensions[col_letter].width = 14

    # =========================================================
    # Sheet 3: Q2 Sales
    # =========================================================
    ws_q2 = wb.create_sheet("Q2 Sales")
    q2_headers = ["Sales Rep", "Region", "April", "May", "June", "Q2 Total"]
    for col, h in enumerate(q2_headers, 1):
        cell = ws_q2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    q2_data = [
        ["Sarah Chen", "Northeast", 49800, 43200, 55600, 148600],
        ["Marcus Johnson", "Southeast", 41200, 45600, 38900, 125700],
        ["Elena Rodriguez", "West Coast", 56300, 52400, 58100, 166800],
        ["David Kim", "Midwest", 34500, 38200, 32600, 105300],
        ["Priya Patel", "Southwest", 47200, 42800, 50100, 140100],
        ["James O'Brien", "Northeast", 40100, 43800, 41500, 125400],
        ["Aisha Williams", "Southeast", 32800, 36400, 39100, 108300],
        ["Carlos Mendez", "West Coast", 51600, 54300, 49200, 155100],
        ["Lisa Chang", "Midwest", 38200, 35600, 41300, 115100],
        ["Robert Taylor", "Southwest", 45300, 47800, 43100, 136200],
        ["Nina Kowalski", "Northeast", 42500, 40200, 46100, 128800],
        ["Ahmed Hassan", "Southeast", 36100, 39500, 34200, 109800],
    ]
    for r, row_data in enumerate(q2_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_q2.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c >= 3:
                cell.number_format = '$#,##0'

    for col_letter in ["A", "B"]:
        ws_q2.column_dimensions[col_letter].width = 18
    for col_letter in ["C", "D", "E", "F"]:
        ws_q2.column_dimensions[col_letter].width = 14

    # =========================================================
    # Sheet 4: Inventory
    # =========================================================
    ws_inv = wb.create_sheet("Inventory")
    inv_headers = ["Product ID", "Product Name", "Category", "Stock Qty", "Reorder Level", "Unit Price"]
    for col, h in enumerate(inv_headers, 1):
        cell = ws_inv.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    inv_data = [
        ["PRD-001", "Wireless Mouse", "Electronics", 245, 50, 29.99],
        ["PRD-002", "USB-C Hub", "Electronics", 132, 30, 49.99],
        ["PRD-003", "Standing Desk Mat", "Furniture", 78, 20, 39.99],
        ["PRD-004", "LED Desk Lamp", "Lighting", 193, 40, 34.99],
        ["PRD-005", "Noise-Cancel Headphones", "Electronics", 67, 25, 149.99],
        ["PRD-006", "Ergonomic Keyboard", "Electronics", 156, 35, 79.99],
        ["PRD-007", "Monitor Arm", "Furniture", 89, 15, 59.99],
        ["PRD-008", "Webcam HD 1080p", "Electronics", 211, 45, 69.99],
        ["PRD-009", "Cable Management Kit", "Accessories", 342, 60, 14.99],
        ["PRD-010", "Laptop Stand", "Furniture", 118, 25, 44.99],
        ["PRD-011", "Wireless Charger", "Electronics", 287, 50, 24.99],
        ["PRD-012", "Desk Organizer", "Accessories", 165, 30, 19.99],
    ]
    for r, row_data in enumerate(inv_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_inv.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 6:
                cell.number_format = '$#,##0.00'

    ws_inv.column_dimensions["A"].width = 12
    ws_inv.column_dimensions["B"].width = 26
    ws_inv.column_dimensions["C"].width = 14
    ws_inv.column_dimensions["D"].width = 12
    ws_inv.column_dimensions["E"].width = 14
    ws_inv.column_dimensions["F"].width = 12

    # =========================================================
    # Sheet 5: Employees
    # =========================================================
    ws_emp = wb.create_sheet("Employees")
    emp_headers = ["Employee ID", "Full Name", "Department", "Title", "Start Date", "Salary"]
    for col, h in enumerate(emp_headers, 1):
        cell = ws_emp.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    emp_data = [
        ["EMP-101", "Sarah Chen", "Engineering", "Senior Developer", "2021-03-15", 125000],
        ["EMP-102", "Marcus Johnson", "Marketing", "Campaign Manager", "2020-06-01", 92000],
        ["EMP-103", "Elena Rodriguez", "Sales", "Regional Director", "2019-01-20", 115000],
        ["EMP-104", "David Kim", "Engineering", "DevOps Engineer", "2022-08-10", 108000],
        ["EMP-105", "Priya Patel", "Product", "Product Manager", "2021-11-05", 118000],
        ["EMP-106", "James O'Brien", "Finance", "Senior Accountant", "2018-04-22", 95000],
        ["EMP-107", "Aisha Williams", "HR", "Recruiting Lead", "2020-09-14", 88000],
        ["EMP-108", "Carlos Mendez", "Engineering", "Staff Engineer", "2017-02-28", 142000],
        ["EMP-109", "Lisa Chang", "Marketing", "Content Strategist", "2023-01-09", 78000],
        ["EMP-110", "Robert Taylor", "Sales", "Account Executive", "2022-05-17", 85000],
        ["EMP-111", "Nina Kowalski", "Engineering", "QA Engineer", "2021-07-30", 96000],
        ["EMP-112", "Ahmed Hassan", "Product", "UX Designer", "2022-12-01", 102000],
    ]
    for r, row_data in enumerate(emp_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_emp.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 6:
                cell.number_format = '$#,##0'

    ws_emp.column_dimensions["A"].width = 14
    ws_emp.column_dimensions["B"].width = 20
    ws_emp.column_dimensions["C"].width = 14
    ws_emp.column_dimensions["D"].width = 20
    ws_emp.column_dimensions["E"].width = 14
    ws_emp.column_dimensions["F"].width = 12

    # Save workbook
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
