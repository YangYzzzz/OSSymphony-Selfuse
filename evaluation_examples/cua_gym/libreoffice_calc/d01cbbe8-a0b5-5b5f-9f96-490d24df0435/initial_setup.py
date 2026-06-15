"""
Initial Setup: Employee list for subtotals task
Task ID: calc_adv_group_subtotals_039
Domain: libreoffice_calc

Creates a spreadsheet with 80 employee records across 5 departments.
Data is NOT sorted by department - mixed order to simulate a real-world scenario.
No subtotals, no grouping, no sorting by department.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_adv_group_subtotals_039'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Employees'

    # Headers
    headers = ['Name', 'Department', 'Role', 'Salary', 'Bonus']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Style header row
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF')
    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Employee data: 80 records, mixed department order (NOT sorted by department)
    # Departments: Engineering(20), Marketing(15), Sales(25), HR(10), Finance(10)
    employees = [
        # Engineering employees (20)
        ('Sarah Chen', 'Engineering', 'Software Engineer', 92000, 9200),
        ('Marcus Johnson', 'Engineering', 'Senior Engineer', 118000, 14160),
        ('Priya Patel', 'Engineering', 'Tech Lead', 135000, 20250),
        ('David Kim', 'Engineering', 'Software Engineer', 88000, 8800),
        ('Emily Rodriguez', 'Engineering', 'DevOps Engineer', 105000, 12600),
        ('James Wilson', 'Engineering', 'QA Engineer', 79000, 7110),
        ('Aisha Mohammed', 'Engineering', 'Data Engineer', 111000, 13320),
        ('Carlos Reyes', 'Engineering', 'Software Engineer', 95000, 9500),
        ('Natalie Foster', 'Engineering', 'Backend Engineer', 102000, 10200),
        ('Ryan Nguyen', 'Engineering', 'Frontend Engineer', 97000, 8730),
        ('Sophia Tanaka', 'Engineering', 'ML Engineer', 125000, 18750),
        ('Tyler Brooks', 'Engineering', 'Software Engineer', 86000, 6880),
        ('Mei Lin', 'Engineering', 'Senior Engineer', 116000, 13920),
        ('Omar Hassan', 'Engineering', 'DevOps Engineer', 108000, 12960),
        ('Laura Bennett', 'Engineering', 'Software Engineer', 91000, 7280),
        ('Kevin Park', 'Engineering', 'Tech Lead', 132000, 19800),
        ('Diana Clark', 'Engineering', 'QA Engineer', 76000, 6080),
        ('Alex Turner', 'Engineering', 'Software Engineer', 89000, 7120),
        ('Nina Zhao', 'Engineering', 'Data Engineer', 113000, 13560),
        ('Brian Scott', 'Engineering', 'Backend Engineer', 98000, 9800),
        # Sales employees (25)
        ('Jennifer Adams', 'Sales', 'Account Executive', 75000, 22500),
        ('Michael Torres', 'Sales', 'Sales Manager', 98000, 34300),
        ('Amanda White', 'Sales', 'Sales Representative', 62000, 12400),
        ('Robert Martinez', 'Sales', 'Account Executive', 78000, 23400),
        ('Linda Thompson', 'Sales', 'Sales Director', 145000, 72500),
        ('Chris Evans', 'Sales', 'Sales Representative', 58000, 11600),
        ('Michelle Lee', 'Sales', 'Account Executive', 81000, 24300),
        ('Daniel Brown', 'Sales', 'Sales Manager', 102000, 35700),
        ('Stephanie Davis', 'Sales', 'Sales Representative', 65000, 13000),
        ('Jonathan Garcia', 'Sales', 'Account Executive', 76000, 22800),
        ('Nicole Anderson', 'Sales', 'Sales Representative', 60000, 12000),
        ('Timothy Jackson', 'Sales', 'Account Manager', 85000, 25500),
        ('Ashley Thomas', 'Sales', 'Sales Representative', 63000, 12600),
        ('Brandon Lewis', 'Sales', 'Account Executive', 79000, 23700),
        ('Rachel Harris', 'Sales', 'Sales Manager', 95000, 33250),
        ('Matthew Robinson', 'Sales', 'Sales Representative', 61000, 12200),
        ('Samantha Walker', 'Sales', 'Account Executive', 77000, 23100),
        ('Joshua Hall', 'Sales', 'Sales Representative', 59000, 11800),
        ('Megan Allen', 'Sales', 'Account Manager', 83000, 24900),
        ('Andrew Young', 'Sales', 'Sales Representative', 64000, 12800),
        ('Crystal Hernandez', 'Sales', 'Account Executive', 80000, 24000),
        ('Nathan King', 'Sales', 'Sales Representative', 62000, 12400),
        ('Brittany Wright', 'Sales', 'Account Manager', 87000, 26100),
        ('Aaron Lopez', 'Sales', 'Sales Representative', 66000, 13200),
        ('Danielle Hill', 'Sales', 'Account Executive', 74000, 22200),
        # Marketing employees (15)
        ('Jessica Scott', 'Marketing', 'Marketing Manager', 89000, 10680),
        ('William Green', 'Marketing', 'Content Strategist', 72000, 7200),
        ('Vanessa Adams', 'Marketing', 'Digital Marketer', 68000, 6800),
        ('Steven Baker', 'Marketing', 'SEO Specialist', 65000, 5850),
        ('Tiffany Nelson', 'Marketing', 'Brand Manager', 85000, 9350),
        ('Gary Carter', 'Marketing', 'Marketing Analyst', 70000, 6300),
        ('Heather Mitchell', 'Marketing', 'Content Creator', 62000, 4960),
        ('Ronald Perez', 'Marketing', 'Marketing Manager', 92000, 11040),
        ('Cheryl Roberts', 'Marketing', 'Social Media Manager', 67000, 5360),
        ('Phillip Turner', 'Marketing', 'Digital Marketer', 71000, 6390),
        ('Jacqueline Phillips', 'Marketing', 'Brand Strategist', 88000, 10560),
        ('Eugene Campbell', 'Marketing', 'Marketing Analyst', 73000, 6570),
        ('Gloria Parker', 'Marketing', 'Content Strategist', 69000, 6210),
        ('Leonard Evans', 'Marketing', 'SEO Specialist', 64000, 5760),
        ('Patricia Edwards', 'Marketing', 'Marketing Director', 118000, 17700),
        # HR employees (10)
        ('Sandra Collins', 'HR', 'HR Manager', 82000, 8200),
        ('Kenneth Stewart', 'HR', 'Recruiter', 63000, 5040),
        ('Dorothy Sanchez', 'HR', 'HR Specialist', 58000, 4060),
        ('Raymond Morris', 'HR', 'Talent Acquisition', 69000, 5520),
        ('Rebecca Rogers', 'HR', 'HR Manager', 85000, 8500),
        ('Frank Reed', 'HR', 'Recruiter', 61000, 4880),
        ('Marie Cook', 'HR', 'HR Specialist', 57000, 3990),
        ('Roger Morgan', 'HR', 'Compensation Analyst', 76000, 6840),
        ('Kathleen Bell', 'HR', 'HR Director', 115000, 17250),
        ('Jerry Murphy', 'HR', 'Training Specialist', 64000, 5120),
        # Finance employees (10)
        ('Ruth Bailey', 'Finance', 'Financial Analyst', 78000, 7800),
        ('Harold Rivera', 'Finance', 'Accountant', 68000, 5440),
        ('Deborah Cooper', 'Finance', 'Finance Manager', 105000, 13650),
        ('Jack Richardson', 'Finance', 'Financial Analyst', 82000, 8200),
        ('Shirley Cox', 'Finance', 'Accountant', 65000, 5200),
        ('Walter Howard', 'Finance', 'Finance Director', 138000, 24840),
        ('Beverly Ward', 'Finance', 'Financial Analyst', 80000, 8000),
        ('Eugene Torres', 'Finance', 'Accountant', 67000, 5360),
        ('Judy Peterson', 'Finance', 'Payroll Specialist', 72000, 5760),
        ('Carl Gray', 'Finance', 'Finance Manager', 108000, 14040),
    ]

    # Write data in a mixed (unsorted) order to simulate real-world scenario
    # Interleave departments so data is NOT sorted by department
    mixed_order = []
    dept_buckets = {
        'Engineering': [],
        'Sales': [],
        'Marketing': [],
        'HR': [],
        'Finance': [],
    }
    for emp in employees:
        dept_buckets[emp[1]].append(emp)

    # Interleave: take from each department in turn to ensure mixed order
    max_len = max(len(v) for v in dept_buckets.values())
    dept_list = ['Engineering', 'Sales', 'Marketing', 'HR', 'Finance']
    for i in range(max_len):
        for dept in dept_list:
            if i < len(dept_buckets[dept]):
                mixed_order.append(dept_buckets[dept][i])

    for r, emp in enumerate(mixed_order, 2):
        ws.cell(row=r, column=1, value=emp[0])  # Name
        ws.cell(row=r, column=2, value=emp[1])  # Department
        ws.cell(row=r, column=3, value=emp[2])  # Role
        ws.cell(row=r, column=4, value=emp[3])  # Salary
        ws.cell(row=r, column=5, value=emp[4])  # Bonus

    # Column widths
    ws.column_dimensions['A'].width = 28  # Name
    ws.column_dimensions['B'].width = 16  # Department
    ws.column_dimensions['C'].width = 28  # Role
    ws.column_dimensions['D'].width = 14  # Salary
    ws.column_dimensions['E'].width = 12  # Bonus

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Employees')
    print(f'  Rows: 81 (1 header + 80 data rows)')
    print(f'  Departments: Engineering(20), Sales(25), Marketing(15), HR(10), Finance(10)')
    print(f'  Data is NOT sorted by department (mixed order)')


create_initial()
