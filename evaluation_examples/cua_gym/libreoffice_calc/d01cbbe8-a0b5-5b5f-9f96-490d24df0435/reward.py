"""
Reward Script: Use Subtotals feature to add department subtotal rows for Salary column
Task ID: calc_adv_group_subtotals_039
Domain: libreoffice_calc
Scoring:
  - Component 1: Data is sorted by Department column (0.25 pts)
  - Component 2: Subtotal rows with SUBTOTAL(9,...) formulas for each of 5 departments (0.40 pts)
  - Component 3: Grand Total row at the bottom of the sheet (0.20 pts)
  - Component 4: Row grouping/outline_level=1 applied to data rows (0.15 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_adv_group_subtotals_039'
EXPECTED_DEPARTMENTS = ['Engineering', 'Finance', 'HR', 'Marketing', 'Sales']  # alphabetical


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Employees' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Employees' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Employees']

    # Gather all data rows (excluding header row 1)
    # Collect department values for all non-subtotal, non-grand-total rows
    all_rows = []
    for row in range(2, ws.max_row + 1):
        vals = [ws.cell(row=row, column=c).value for c in range(1, 6)]
        all_rows.append((row, vals))

    # -------------------------------------------------------------------------
    # Component 1: Data is sorted by Department (column B) (0.25 points)
    # The initial file has unsorted data; sorting must have been applied.
    # We verify by collecting non-subtotal/non-grand-total employee rows
    # and checking that Department column (B) is in non-decreasing order.
    # -------------------------------------------------------------------------
    try:
        # Collect only data rows (skip subtotal rows and grand total)
        # A subtotal row typically has a SUBTOTAL formula in column D or has
        # a name like "<Dept> Sum" in column A. A data row has a numeric salary.
        data_dept_sequence = []
        for row_num, vals in all_rows:
            name_val = vals[0]   # A: Name
            dept_val = vals[1]   # B: Department
            salary_val = vals[3]  # D: Salary
            # Skip rows that are subtotal rows (formula in D) or grand total
            if isinstance(salary_val, str) and 'SUBTOTAL' in str(salary_val).upper():
                continue
            if name_val == 'Grand Total':
                continue
            # Also skip rows where name ends with ' Sum' (department subtotal label)
            if isinstance(name_val, str) and name_val.endswith(' Sum'):
                continue
            if dept_val is not None:
                data_dept_sequence.append(dept_val)

        # Check that departments appear in sorted (grouped) order:
        # All records for one dept should be contiguous and alphabetically ordered
        if len(data_dept_sequence) == 80:
            # Build the actual unique-dept sequence in appearance order
            dept_order = []
            for d in data_dept_sequence:
                if not dept_order or dept_order[-1] != d:
                    dept_order.append(d)
            # Must be alphabetically sorted (all 5 departments in order)
            is_sorted = dept_order == sorted(dept_order)
            all_depts_present = sorted(dept_order) == EXPECTED_DEPARTMENTS
            if is_sorted and all_depts_present:
                print(f"PASS: Component 1 — Data sorted by Department, dept order: {dept_order} (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 1 — Data not properly sorted by Department. Found order: {dept_order}")
        else:
            print(f"FAIL: Component 1 — Expected 80 employee data rows, found {len(data_dept_sequence)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Subtotal rows with SUBTOTAL(9,...) formulas for each dept (0.40 pts)
    # Expectation: 5 department subtotal rows in the sheet, each having a
    # SUBTOTAL formula in column D, appearing after the last employee row for
    # that department. Score 0.08 per correct subtotal row (5 depts x 0.08 = 0.40).
    # -------------------------------------------------------------------------
    try:
        subtotal_rows_found = {}
        for row_num, vals in all_rows:
            name_val = vals[0]
            dept_val = vals[1]
            salary_formula = vals[3]
            # Identify subtotal rows: name is "<Dept> Sum" and D contains SUBTOTAL formula
            if (isinstance(name_val, str) and name_val.endswith(' Sum') and
                    isinstance(salary_formula, str) and
                    'SUBTOTAL' in salary_formula.upper()):
                dept_name = name_val.replace(' Sum', '')
                subtotal_rows_found[dept_name] = (row_num, salary_formula)

        pts_per_dept = 0.08
        for dept in EXPECTED_DEPARTMENTS:
            if dept in subtotal_rows_found:
                row_num, formula = subtotal_rows_found[dept]
                print(f"PASS: Component 2 [{dept}] — Subtotal row found at row {row_num}, formula: {formula} ({pts_per_dept} pts)")
                total_score += pts_per_dept
            else:
                print(f"FAIL: Component 2 [{dept}] — No subtotal row found with SUBTOTAL formula")

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Grand Total row at the bottom with SUBTOTAL formula (0.20 points)
    # Expectation: Last row has name 'Grand Total' in column A and a
    # SUBTOTAL formula in column D.
    # -------------------------------------------------------------------------
    try:
        last_row_num = ws.max_row
        last_row_vals = [ws.cell(row=last_row_num, column=c).value for c in range(1, 6)]
        name_val = last_row_vals[0]
        salary_formula = last_row_vals[3]

        if (name_val == 'Grand Total' and
                isinstance(salary_formula, str) and
                'SUBTOTAL' in salary_formula.upper()):
            print(f"PASS: Component 3 — Grand Total row at row {last_row_num}, formula: {salary_formula} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 3 — Expected Grand Total row at row {last_row_num}, found: name={name_val!r}, formula={salary_formula!r}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: Row grouping/outline applied to data rows (0.15 points)
    # The LibreOffice Subtotals feature creates row groups allowing
    # collapse/expand. This is stored as outline_level >= 1 on data rows.
    # Expectation: All 80 employee data rows have outline_level=1.
    # -------------------------------------------------------------------------
    try:
        grouped_data_rows = 0
        # Identify data rows (not subtotal, not grand total, not header)
        for row_num, vals in all_rows:
            name_val = vals[0]
            salary_val = vals[3]
            # Skip subtotal rows (formula in D) and grand total
            if isinstance(salary_val, str) and 'SUBTOTAL' in str(salary_val).upper():
                continue
            if name_val == 'Grand Total':
                continue
            if isinstance(name_val, str) and name_val.endswith(' Sum'):
                continue
            # Check outline_level on this row
            rd = ws.row_dimensions.get(row_num)
            if rd and rd.outline_level and rd.outline_level >= 1:
                grouped_data_rows += 1

        if grouped_data_rows >= 80:
            print(f"PASS: Component 4 — {grouped_data_rows} data rows have row grouping (outline_level >= 1) (0.15 pts)")
            total_score += 0.15
        elif grouped_data_rows >= 40:
            # Partial credit: at least half the rows grouped
            partial = 0.07
            print(f"PARTIAL: Component 4 — {grouped_data_rows}/80 data rows have row grouping; partial credit ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {grouped_data_rows}/80 data rows have row grouping (outline_level >= 1)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
