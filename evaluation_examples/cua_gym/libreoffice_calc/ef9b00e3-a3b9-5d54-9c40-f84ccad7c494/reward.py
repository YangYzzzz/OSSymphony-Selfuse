"""
Reward Script: FormatAsTable macro formatting verification
Task ID: calc_mcp_010
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): Bold headers in row 1 (A1:F1)
  Component 2 (0.4): Alternating row colors rows 2-20 (even=light blue, odd=white)
  Component 3 (0.3): Thin borders on all cells A1:F20
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_010'

# Expected colors (ARGB 8-char)
LIGHT_BLUE = 'FFDCE6F1'
WHITE = 'FFFFFFFF'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the Report sheet exists
    if 'Report' not in wb.sheetnames:
        print("CRITICAL: 'Report' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Report']

    # Component 1: Bold headers in row 1, A1:F1 (0.3 points)
    # This should FAIL on initial (no bold) and PASS on golden (bold applied)
    try:
        bold_count = 0
        total_headers = 6
        for col in range(1, 7):
            cell = ws.cell(row=1, column=col)
            if cell.font.bold:
                bold_count += 1
            else:
                print(f"FAIL: Component 1 — {cell.coordinate} is not bold (font.bold={cell.font.bold})")

        if bold_count == total_headers:
            print(f"PASS: Component 1 — All {total_headers} headers in row 1 are bold (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — Only {bold_count}/{total_headers} headers are bold")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Alternating row colors for rows 2-20 (0.4 points)
    # Even rows (2,4,6,...,20) should be light blue, odd rows (3,5,7,...,19) should be white
    # This should FAIL on initial (no fill) and PASS on golden (alternating fills)
    try:
        correct_colors = 0
        total_data_rows = 19  # rows 2-20
        for row in range(2, 21):
            cell = ws.cell(row=row, column=1)
            fill_type = cell.fill.fill_type
            fg_rgb = None
            try:
                fg_rgb = cell.fill.fgColor.rgb
            except Exception:
                pass

            if row % 2 == 0:
                # Even rows should be light blue
                expected = LIGHT_BLUE
            else:
                # Odd rows should be white
                expected = WHITE

            if fill_type == 'solid' and fg_rgb == expected:
                correct_colors += 1
            else:
                print(f"FAIL: Component 2 — Row {row}: expected fill={expected}, "
                      f"got fill_type={fill_type}, fg={fg_rgb}")

        if correct_colors == total_data_rows:
            print(f"PASS: Component 2 — All {total_data_rows} data rows have correct alternating colors (0.4 pts)")
            total_score += 0.4
        elif correct_colors >= total_data_rows * 0.8:
            partial = round(0.4 * (correct_colors / total_data_rows), 2)
            if partial > 0:
                print(f"PARTIAL: Component 2 — {correct_colors}/{total_data_rows} rows correct, awarding {partial} pts")
                total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {correct_colors}/{total_data_rows} rows have correct colors")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Thin borders on all cells A1:F20 (0.3 points)
    # All cells in range should have thin borders on all 4 sides
    # This should FAIL on initial (no borders) and PASS on golden (thin borders)
    try:
        bordered_cells = 0
        total_cells = 20 * 6  # 20 rows x 6 columns = 120 cells
        for row in range(1, 21):
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                b = cell.border
                if (b.left.style == 'thin' and b.right.style == 'thin' and
                        b.top.style == 'thin' and b.bottom.style == 'thin'):
                    bordered_cells += 1

        if bordered_cells == total_cells:
            print(f"PASS: Component 3 — All {total_cells} cells have thin borders (0.3 pts)")
            total_score += 0.3
        elif bordered_cells >= total_cells * 0.8:
            partial = round(0.3 * (bordered_cells / total_cells), 2)
            if partial > 0:
                print(f"PARTIAL: Component 3 — {bordered_cells}/{total_cells} cells bordered, awarding {partial} pts")
                total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {bordered_cells}/{total_cells} cells have thin borders")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
