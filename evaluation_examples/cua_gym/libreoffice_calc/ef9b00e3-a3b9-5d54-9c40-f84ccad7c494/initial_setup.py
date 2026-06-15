"""
Initial Setup: Create a spreadsheet with employee data, no formatting or macros.
Task ID: calc_mcp_010
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_010'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Report'

    # Headers
    headers = ['Employee ID', 'Name', 'Department', 'Salary', 'Start Date', 'Status']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 19 rows of realistic employee data (rows 2-20)
    data = [
        ['EMP001', 'Sarah Chen', 'Engineering', 95000, '2022-03-15', 'Active'],
        ['EMP002', 'Marcus Johnson', 'Marketing', 72000, '2021-08-22', 'Active'],
        ['EMP003', 'Priya Patel', 'Finance', 88000, '2020-11-03', 'Active'],
        ['EMP004', 'James O\'Brien', 'Engineering', 102000, '2019-06-18', 'Active'],
        ['EMP005', 'Mei-Lin Wu', 'Human Resources', 67000, '2023-01-10', 'Active'],
        ['EMP006', 'Carlos Rivera', 'Sales', 78000, '2022-09-05', 'On Leave'],
        ['EMP007', 'Aisha Mohammed', 'Engineering', 91000, '2021-04-14', 'Active'],
        ['EMP008', 'Daniel Kim', 'Finance', 83000, '2020-07-28', 'Active'],
        ['EMP009', 'Emma Larsson', 'Marketing', 69000, '2023-05-20', 'Active'],
        ['EMP010', 'Raj Gupta', 'Engineering', 97000, '2019-12-01', 'Active'],
        ['EMP011', 'Olivia Thompson', 'Sales', 74000, '2022-02-14', 'On Leave'],
        ['EMP012', 'Kenji Tanaka', 'Engineering', 105000, '2018-10-30', 'Active'],
        ['EMP013', 'Sofia Andersson', 'Human Resources', 71000, '2021-11-08', 'Active'],
        ['EMP014', 'William Okafor', 'Finance', 86000, '2020-03-25', 'Active'],
        ['EMP015', 'Hannah Mueller', 'Marketing', 68000, '2023-07-12', 'Probation'],
        ['EMP016', 'Lucas Ferreira', 'Sales', 81000, '2021-01-19', 'Active'],
        ['EMP017', 'Fatima Al-Rashid', 'Engineering', 99000, '2020-05-06', 'Active'],
        ['EMP018', 'Ryan Mitchell', 'Finance', 77000, '2022-08-15', 'Active'],
        ['EMP019', 'Yuki Nakamura', 'Human Resources', 65000, '2023-09-01', 'Probation'],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Open in LibreOffice Calc for GUI-ready state
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
