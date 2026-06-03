"""
Initial Setup: Create production_data.xlsx with monthly production and efficiency data
Task ID: calc_gg5_018
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_018'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Output ---
    ws = wb.active
    ws.title = 'Output'

    # Headers
    headers = ['Month', 'Production Volume', 'Efficiency Rate']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")

    # Monthly production data (12 months)
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    production_volumes = [
        78500, 82300, 91200, 105600, 112400, 134800,
        142500, 138700, 125300, 108900, 95600, 68200
    ]
    efficiency_rates = [
        0.82, 0.84, 0.87, 0.91, 0.93, 0.95,
        0.94, 0.92, 0.89, 0.86, 0.83, 0.78
    ]

    for r, (month, vol, eff) in enumerate(zip(months, production_volumes, efficiency_rates), 2):
        ws.cell(row=r, column=1, value=month)
        cell_vol = ws.cell(row=r, column=2, value=vol)
        cell_vol.number_format = '#,##0'
        cell_eff = ws.cell(row=r, column=3, value=eff)
        cell_eff.number_format = '0.00%'

    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18

    # NO charts in initial state

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
