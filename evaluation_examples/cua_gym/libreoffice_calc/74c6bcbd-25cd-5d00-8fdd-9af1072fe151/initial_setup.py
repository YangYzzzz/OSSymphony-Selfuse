"""
Initial Setup: Equipment Booking Tracker for Forklifts and Pallet Jacks
Task ID: calc_ops_resource_equipment_booking_037
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import datetime

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_resource_equipment_booking_037'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'EquipmentBookings'

    # Headers (Row 1)
    headers = [
        'Booking ID', 'Equipment ID', 'Equipment Type', 'Booked By',
        'Date', 'Start Time', 'End Time', 'Purpose', 'Double Book Flag'
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Equipment options
    equipment_list = [
        ('FL-001', 'Forklift'),
        ('FL-002', 'Forklift'),
        ('FL-003', 'Forklift'),
        ('PJ-001', 'Pallet Jack'),
        ('PJ-002', 'Pallet Jack'),
    ]

    # Staff names
    staff = [
        'Carlos Mendez', 'Aisha Okonkwo', 'Derek Lam', 'Priya Nair',
        'Tomasz Wiecki', 'Fatima Al-Hassan', 'James O\'Brien', 'Mei-Ling Zhao',
        'Rodrigo Santos', 'Elaine Kowalski', 'Bjorn Eriksen', 'Nadia Petrov',
        'Samuel Abebe', 'Laura Hoffmann', 'Kevin Tran'
    ]

    # Purposes
    purposes = [
        'Unloading delivery truck', 'Moving pallet to Warehouse B',
        'Loading outbound shipment', 'Restocking shelves in Zone 3',
        'Transferring heavy equipment', 'End-of-day pallet organization',
        'Receiving dock clearance', 'Cross-docking freight',
        'Seasonal inventory rotation', 'Emergency restock',
        'Pre-shift area setup', 'Monthly stocktake support',
        'Customer order fulfillment', 'Bulk material handling',
        'Inbound freight sorting'
    ]

    # Generate booking data — 70 rows
    # Dates range: Jan to Mar 2025
    # NOTE: DO NOT sort data — it will be unsorted in the initial file
    # Some bookings intentionally have same Equipment ID + same Date
    # (to create double-booking scenarios for the formula to detect)

    import random
    random.seed(42)  # reproducible

    booking_data = []

    # Base dates: 30 unique workdays in Jan-Mar 2025
    base_dates = []
    d = datetime.date(2025, 1, 6)  # Monday
    while len(base_dates) < 30 and d <= datetime.date(2025, 3, 31):
        if d.weekday() < 5:  # Mon-Fri
            base_dates.append(d)
        d += datetime.timedelta(days=1)

    # Time slots: 8 possible start times
    start_times = [
        datetime.time(7, 0),
        datetime.time(8, 0),
        datetime.time(9, 30),
        datetime.time(10, 0),
        datetime.time(11, 30),
        datetime.time(13, 0),
        datetime.time(14, 30),
        datetime.time(16, 0),
    ]
    duration_hours = 1  # each slot is 1 hour

    # Create 70 bookings with some intentional double bookings
    booking_id_counter = 1001
    used_slots = {}  # track (equip_id, date, start_time) to inject double bookings

    for i in range(70):
        equip_id, equip_type = random.choice(equipment_list)
        date = random.choice(base_dates)
        start_t = random.choice(start_times)
        end_t = datetime.time(start_t.hour + duration_hours, start_t.minute)
        booked_by = random.choice(staff)
        purpose = random.choice(purposes)
        booking_id = f'BK-{booking_id_counter}'
        booking_id_counter += 1

        booking_data.append({
            'booking_id': booking_id,
            'equip_id': equip_id,
            'equip_type': equip_type,
            'booked_by': booked_by,
            'date': date,
            'start_time': start_t,
            'end_time': end_t,
            'purpose': purpose,
        })

    # Write data rows (NOT sorted — agent must sort)
    for row_idx, entry in enumerate(booking_data, 2):
        ws.cell(row=row_idx, column=1, value=entry['booking_id'])
        ws.cell(row=row_idx, column=2, value=entry['equip_id'])
        ws.cell(row=row_idx, column=3, value=entry['equip_type'])
        ws.cell(row=row_idx, column=4, value=entry['booked_by'])

        # Date as date value
        date_cell = ws.cell(row=row_idx, column=5, value=entry['date'])
        date_cell.number_format = 'yyyy-mm-dd'

        # Start time as time value
        start_cell = ws.cell(row=row_idx, column=6, value=entry['start_time'])
        start_cell.number_format = 'HH:MM'

        # End time as time value
        end_cell = ws.cell(row=row_idx, column=7, value=entry['end_time'])
        end_cell.number_format = 'HH:MM'

        ws.cell(row=row_idx, column=8, value=entry['purpose'])
        # Column I (Double Book Flag) is EMPTY — agent fills this

    # Set column widths
    col_widths = {
        'A': 12,  # Booking ID
        'B': 12,  # Equipment ID
        'C': 14,  # Equipment Type
        'D': 20,  # Booked By
        'E': 14,  # Date
        'F': 12,  # Start Time
        'G': 12,  # End Time
        'H': 30,  # Purpose
        'I': 18,  # Double Book Flag
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Rows: 70 booking entries (rows 2-71)')
    print(f'  Column I (Double Book Flag): EMPTY — no formulas')
    print(f'  Data: NOT sorted, NO data validation, NO conditional formatting')


create_initial()
