"""
Reward Script: Equipment Booking Tracker with COUNTIFS Double-Booking Detection
Task ID: calc_ops_resource_equipment_booking_037
Domain: libreoffice_calc
Scoring:
  Component 1: COUNTIFS formula in I2:I71 for double-booking detection (0.40 pts)
  Component 2: Data validation dropdown on B2:B71 for Equipment IDs (0.20 pts)
  Component 3: Conditional formatting on I column with red fill for DOUBLE BOOKED (0.20 pts)
  Component 4: Data sorted by Equipment ID then Date then Start Time (0.20 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_ops_resource_equipment_booking_037'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get sheet
    sheet_name = 'EquipmentBookings'
    if sheet_name not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{sheet_name}' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[sheet_name]

    # Component 1: COUNTIFS formula in I2:I71 for double-booking detection (0.40 pts)
    # The formula should be =IF(COUNTIFS($B:$B,Bx,$E:$E,Ex)>1,"DOUBLE BOOKED","")
    # This FAILS on initial (all None) → PASSES on golden (all have formula)
    try:
        rows_with_formula = 0
        total_data_rows = 70  # rows 2 to 71
        formula_examples = []
        for row in range(2, 72):
            cell_val = ws.cell(row=row, column=9).value  # Column I
            if cell_val is not None and isinstance(cell_val, str) and cell_val.startswith('='):
                # Check it's a COUNTIFS-based formula referencing columns B and E
                val_upper = cell_val.upper()
                if 'COUNTIFS' in val_upper and '$B:$B' in val_upper and '$E:$E' in val_upper:
                    rows_with_formula += 1
                    if len(formula_examples) < 2:
                        formula_examples.append(f"I{row}: {cell_val}")

        if rows_with_formula == total_data_rows:
            print(f"PASS: Component 1 — COUNTIFS formula in all {total_data_rows} rows of column I (0.40 pts)")
            print(f"      Example: {formula_examples[0] if formula_examples else 'N/A'}")
            total_score += 0.40
        elif rows_with_formula > 0:
            partial = round(0.40 * rows_with_formula / total_data_rows, 4)
            print(f"PARTIAL: Component 1 — COUNTIFS formula in {rows_with_formula}/{total_data_rows} rows of column I ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No COUNTIFS formula found in column I (expected in I2:I71)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Data validation dropdown on B2:B71 for Equipment IDs (0.20 pts)
    # This FAILS on initial (no data validation) → PASSES on golden (has DV with equipment IDs)
    try:
        dvs = ws.data_validations.dataValidation
        dv_found = False
        for dv in dvs:
            if dv.type == 'list' and dv.formula1 is not None:
                # Check formula contains expected equipment IDs
                formula = dv.formula1
                # Check sqref covers B2:B71
                sqref_str = str(dv.sqref)
                if 'B2' in sqref_str and 'B71' in sqref_str:
                    # Verify equipment IDs are present in the formula
                    required_ids = ['FL-001', 'FL-002', 'FL-003', 'PJ-001', 'PJ-002']
                    ids_found = all(eq_id in formula for eq_id in required_ids)
                    if ids_found:
                        print(f"PASS: Component 2 — Data validation dropdown on B2:B71 with all equipment IDs (0.20 pts)")
                        print(f"      formula1={formula}, sqref={sqref_str}")
                        dv_found = True
                        total_score += 0.20
                        break

        if not dv_found:
            # Check if any list validation exists at all (partial credit scenario)
            any_list_dv = False
            for dv in dvs:
                if dv.type == 'list' and dv.formula1 is not None:
                    formula = dv.formula1
                    if any(eq_id in formula for eq_id in ['FL-001', 'FL-002', 'FL-003', 'PJ-001', 'PJ-002']):
                        any_list_dv = True
                        print(f"FAIL: Component 2 — Data validation found but wrong range. formula1={formula}, sqref={dv.sqref} (expected B2:B71)")
                        break
            if not any_list_dv:
                print(f"FAIL: Component 2 — No data validation dropdown found on B2:B71 for equipment IDs")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Conditional formatting on I2:I71 with red fill for "DOUBLE BOOKED" (0.20 pts)
    # This FAILS on initial (no CF) → PASSES on golden (has CF rule with red fill)
    try:
        cf_found = False
        for cf_range, cf_rules in ws.conditional_formatting._cf_rules.items():
            cf_range_str = str(cf_range)
            # Check it covers column I
            if 'I' in cf_range_str:
                for rule in cf_rules:
                    # Check for expression rule with "DOUBLE BOOKED"
                    has_double_book_formula = False
                    if hasattr(rule, 'formula') and rule.formula:
                        for f in rule.formula:
                            if 'DOUBLE BOOKED' in str(f).upper():
                                has_double_book_formula = True
                                break
                    # Check for red fill
                    has_red_fill = False
                    if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                        try:
                            fg_color = rule.dxf.fill.fgColor.rgb
                            # Red fill: FFFF0000 or close variants
                            if fg_color and 'FF0000' in fg_color.upper():
                                has_red_fill = True
                        except Exception:
                            pass

                    if has_double_book_formula and has_red_fill:
                        print(f"PASS: Component 3 — Conditional formatting on column I with red fill for 'DOUBLE BOOKED' (0.20 pts)")
                        print(f"      Range: {cf_range_str}")
                        cf_found = True
                        total_score += 0.20
                        break
                if cf_found:
                    break

        if not cf_found:
            # Check if there is any CF on column I (even without red fill)
            any_cf_on_i = any('I' in str(r) for r in ws.conditional_formatting._cf_rules.keys())
            if any_cf_on_i:
                print(f"FAIL: Component 3 — Conditional formatting on column I found but missing red fill or 'DOUBLE BOOKED' formula")
            else:
                print(f"FAIL: Component 3 — No conditional formatting found on column I (I2:I71)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Data sorted by Equipment ID (col B) then Date (col E) then Start Time (col F) (0.20 pts)
    # This FAILS on initial (data is not sorted) → PASSES on golden (data is sorted)
    try:
        sort_violations = 0
        prev_b = prev_e = prev_f = None
        for r in range(2, 72):
            b = ws.cell(row=r, column=2).value   # Equipment ID
            e = ws.cell(row=r, column=5).value   # Date
            f = ws.cell(row=r, column=6).value   # Start Time

            if prev_b is not None and b is not None and e is not None:
                curr_tuple = (str(b) if b else '', e, f)
                prev_tuple = (str(prev_b) if prev_b else '', prev_e, prev_f)
                try:
                    if curr_tuple < prev_tuple:
                        sort_violations += 1
                except TypeError:
                    # Handle incomparable types by comparing string representations
                    pass
            prev_b, prev_e, prev_f = b, e, f

        if sort_violations == 0:
            print(f"PASS: Component 4 — Data sorted by Equipment ID, Date, Start Time (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 4 — Data is NOT sorted correctly ({sort_violations} sort violations found)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
