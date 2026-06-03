"""
Initial Setup: Set up data validation on cell D2 using a list from Lookup sheet
Task ID: calc_nrv_062
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_062'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Main sheet ---
    ws_main = wb.active
    ws_main.title = 'Main'

    # Headers
    headers = ['Employee', 'Hours', 'Date', 'Project Code']
    for col, h in enumerate(headers, 1):
        ws_main.cell(row=1, column=col, value=h)

    # Realistic employee data (D column has project codes for rows 3+ but D2 is EMPTY)
    data = [
        ['Sarah Chen',       38.5, '2025-03-10', None],          # D2 empty - this is where validation goes
        ['Marcus Johnson',   42.0, '2025-03-10', 'PRJ-005'],
        ['Elena Rodriguez',  36.0, '2025-03-11', 'PRJ-012'],
        ['James O\'Brien',   40.0, '2025-03-11', 'PRJ-003'],
        ['Aisha Patel',      37.5, '2025-03-12', 'PRJ-018'],
        ['David Kim',        41.0, '2025-03-12', 'PRJ-007'],
        ['Olivia Thompson',  39.0, '2025-03-13', 'PRJ-021'],
        ['Wei Zhang',        44.0, '2025-03-13', 'PRJ-002'],
        ['Rachel Foster',    35.5, '2025-03-14', 'PRJ-009'],
        ['Carlos Mendez',    40.0, '2025-03-14', 'PRJ-015'],
        ['Hannah Brooks',    38.0, '2025-03-15', 'PRJ-026'],
        ['Liam O\'Connor',   42.5, '2025-03-15', 'PRJ-011'],
        ['Priya Sharma',     36.5, '2025-03-16', 'PRJ-004'],
        ['Nathan Wright',    39.5, '2025-03-16', 'PRJ-022'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            if val is not None:
                ws_main.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws_main.column_dimensions['A'].width = 20
    ws_main.column_dimensions['B'].width = 10
    ws_main.column_dimensions['C'].width = 14
    ws_main.column_dimensions['D'].width = 14

    # --- Lookup sheet ---
    ws_lookup = wb.create_sheet('Lookup')
    ws_lookup.cell(row=1, column=1, value='Project Codes')

    # A2:A30 with project codes PRJ-001 through PRJ-029
    for i in range(1, 30):
        ws_lookup.cell(row=i + 1, column=1, value=f'PRJ-{i:03d}')

    ws_lookup.column_dimensions['A'].width = 15

    # NO validation on D2 - that is the task for the agent

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
