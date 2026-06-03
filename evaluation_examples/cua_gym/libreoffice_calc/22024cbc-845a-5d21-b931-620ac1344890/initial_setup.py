"""
Initial Setup: Format cells D2:D25 as dates in DD-MMM-YYYY format
Task ID: calc_gfl_031
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
from datetime import datetime, timedelta
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_031'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Schedule'

    # Headers
    headers = ['Task ID', 'Task Name', 'Owner', 'Start Date', 'End Date', 'Duration', 'Status']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Project task data - 24 rows of realistic project management data
    tasks = [
        ['T-001', 'Requirements Gathering', 'Sarah Chen', datetime(2024, 1, 15), datetime(2024, 1, 26), 10, 'Completed'],
        ['T-002', 'Stakeholder Interviews', 'Marcus Johnson', datetime(2024, 1, 18), datetime(2024, 1, 31), 10, 'Completed'],
        ['T-003', 'Market Research Analysis', 'Priya Patel', datetime(2024, 2, 1), datetime(2024, 2, 14), 10, 'Completed'],
        ['T-004', 'Technical Feasibility Study', 'David Kim', datetime(2024, 2, 5), datetime(2024, 2, 23), 15, 'Completed'],
        ['T-005', 'Architecture Design', 'Elena Rodriguez', datetime(2024, 2, 12), datetime(2024, 3, 1), 14, 'Completed'],
        ['T-006', 'Database Schema Design', 'James Wright', datetime(2024, 2, 19), datetime(2024, 3, 8), 14, 'Completed'],
        ['T-007', 'UI/UX Wireframes', 'Aisha Mohammed', datetime(2024, 3, 1), datetime(2024, 3, 15), 11, 'Completed'],
        ['T-008', 'API Specification', 'Carlos Mendez', datetime(2024, 3, 4), datetime(2024, 3, 22), 15, 'Completed'],
        ['T-009', 'Sprint 1 - Core Backend', 'Sarah Chen', datetime(2024, 3, 11), datetime(2024, 3, 29), 15, 'Completed'],
        ['T-010', 'Sprint 1 - Frontend Setup', 'David Kim', datetime(2024, 3, 18), datetime(2024, 4, 5), 15, 'In Progress'],
        ['T-011', 'Sprint 2 - Auth Module', 'Elena Rodriguez', datetime(2024, 4, 1), datetime(2024, 4, 19), 15, 'In Progress'],
        ['T-012', 'Sprint 2 - Dashboard', 'Priya Patel', datetime(2024, 4, 8), datetime(2024, 4, 26), 15, 'In Progress'],
        ['T-013', 'Integration Testing Phase 1', 'Marcus Johnson', datetime(2024, 4, 15), datetime(2024, 4, 30), 12, 'Not Started'],
        ['T-014', 'Security Audit', 'James Wright', datetime(2024, 5, 1), datetime(2024, 5, 17), 13, 'Not Started'],
        ['T-015', 'Performance Optimization', 'Aisha Mohammed', datetime(2024, 5, 6), datetime(2024, 5, 24), 15, 'Not Started'],
        ['T-016', 'Sprint 3 - Reporting Module', 'Carlos Mendez', datetime(2024, 5, 13), datetime(2024, 5, 31), 15, 'Not Started'],
        ['T-017', 'Sprint 3 - Notifications', 'Sarah Chen', datetime(2024, 5, 20), datetime(2024, 6, 7), 15, 'Not Started'],
        ['T-018', 'User Acceptance Testing', 'David Kim', datetime(2024, 6, 3), datetime(2024, 6, 21), 15, 'Not Started'],
        ['T-019', 'Bug Fixes - Round 1', 'Elena Rodriguez', datetime(2024, 6, 10), datetime(2024, 6, 28), 15, 'Not Started'],
        ['T-020', 'Documentation', 'Priya Patel', datetime(2024, 6, 17), datetime(2024, 7, 5), 15, 'Not Started'],
        ['T-021', 'Deployment Preparation', 'Marcus Johnson', datetime(2024, 7, 1), datetime(2024, 7, 12), 10, 'Not Started'],
        ['T-022', 'Staging Environment Setup', 'James Wright', datetime(2024, 7, 8), datetime(2024, 7, 19), 10, 'Not Started'],
        ['T-023', 'Production Deployment', 'Aisha Mohammed', datetime(2024, 7, 15), datetime(2024, 7, 26), 10, 'Not Started'],
        ['T-024', 'Post-Launch Monitoring', 'Carlos Mendez', datetime(2024, 7, 22), datetime(2024, 8, 2), 10, 'Not Started'],
    ]

    for r, row_data in enumerate(tasks, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 14

    # Leave dates in default format (no custom DD-MMM-YYYY)
    # The task is for the agent to apply that format

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
