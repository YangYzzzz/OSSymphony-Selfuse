"""
Reward Script: Format cells D2:D25 as dates in DD-MMM-YYYY format
Task ID: calc_gfl_031
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): At least 50% of D2:D25 use a DD-MMM date format variant
  Component 2 (0.4): ALL 24 cells D2:D25 use exactly DD-MMM-YYYY format
  Component 3 (0.2): Format is DD-MMM-YYYY AND underlying date values preserved
"""

import os
import datetime

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_031'

# Expected date values in D2:D25 (from the task data)
EXPECTED_DATES = [
    datetime.datetime(2024, 1, 15),
    datetime.datetime(2024, 1, 18),
    datetime.datetime(2024, 2, 1),
    datetime.datetime(2024, 2, 5),
    datetime.datetime(2024, 2, 12),
    datetime.datetime(2024, 2, 19),
    datetime.datetime(2024, 3, 1),
    datetime.datetime(2024, 3, 4),
    datetime.datetime(2024, 3, 11),
    datetime.datetime(2024, 3, 18),
    datetime.datetime(2024, 4, 1),
    datetime.datetime(2024, 4, 8),
    datetime.datetime(2024, 4, 15),
    datetime.datetime(2024, 5, 1),
    datetime.datetime(2024, 5, 6),
    datetime.datetime(2024, 5, 13),
    datetime.datetime(2024, 5, 20),
    datetime.datetime(2024, 6, 3),
    datetime.datetime(2024, 6, 10),
    datetime.datetime(2024, 6, 17),
    datetime.datetime(2024, 7, 1),
    datetime.datetime(2024, 7, 8),
    datetime.datetime(2024, 7, 15),
    datetime.datetime(2024, 7, 22),
]


def is_dd_mmm_format(fmt):
    """Check if a number_format string matches DD-MMM-YYYY or close variants."""
    if fmt is None:
        return False
    fmt_upper = fmt.upper().strip()
    # Accept common DD-MMM-YYYY variants (case-insensitive, with or without extra chars)
    # Patterns: DD-MMM-YYYY, DD-MMM-YY, D-MMM-YYYY, DD/MMM/YYYY etc.
    acceptable = [
        'DD-MMM-YYYY',
        'DD-MMM-YY',
        'D-MMM-YYYY',
        'D-MMM-YY',
        'DD/MMM/YYYY',
        'DD/MMM/YY',
    ]
    return fmt_upper in acceptable


def is_dd_mmm_yyyy_exact(fmt):
    """Check for the exact DD-MMM-YYYY format."""
    if fmt is None:
        return False
    return fmt.upper().strip() == 'DD-MMM-YYYY'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sheet 'Schedule' must exist
    if 'Schedule' not in wb.sheetnames:
        print(f"FAIL: Sheet 'Schedule' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Schedule']

    # Component 1: At least 50% of D2:D25 use a DD-MMM date format variant (0.4 points)
    # This checks that the formatting change was broadly applied.
    try:
        dd_mmm_count = 0
        total_cells = 24  # D2 through D25
        for r in range(2, 26):
            cell = ws.cell(row=r, column=4)
            if is_dd_mmm_format(cell.number_format):
                dd_mmm_count += 1
        pct = dd_mmm_count / total_cells
        if pct >= 0.5:
            print(f"PASS: Component 1 — {dd_mmm_count}/{total_cells} cells ({pct*100:.0f}%) have DD-MMM format (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — only {dd_mmm_count}/{total_cells} cells ({pct*100:.0f}%) have DD-MMM format, need >=50%")
            # Show a few samples for debugging
            for r in [2, 3, 4]:
                cell = ws.cell(row=r, column=4)
                print(f"  D{r}: number_format={cell.number_format!r}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: ALL 24 cells D2:D25 have exactly DD-MMM-YYYY format (0.4 points)
    # This checks full, precise coverage.
    try:
        exact_count = 0
        for r in range(2, 26):
            cell = ws.cell(row=r, column=4)
            if is_dd_mmm_yyyy_exact(cell.number_format):
                exact_count += 1
        if exact_count == total_cells:
            print(f"PASS: Component 2 — all {total_cells} cells have exact DD-MMM-YYYY format (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 2 — {exact_count}/{total_cells} cells have exact DD-MMM-YYYY format")
            # Show mismatches
            for r in range(2, 26):
                cell = ws.cell(row=r, column=4)
                if not is_dd_mmm_yyyy_exact(cell.number_format):
                    print(f"  D{r}: number_format={cell.number_format!r} (expected DD-MMM-YYYY)")
                    break  # show only first mismatch
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Format is DD-MMM-YYYY AND underlying date values are preserved (0.2 points)
    # This is a compound check: format change was applied AND data integrity is maintained.
    # Both conditions must be true — fails on initial because format is not DD-MMM-YYYY.
    try:
        format_fail_count = 0
        value_fail_count = 0
        mismatches = []
        for i, r in enumerate(range(2, 26)):
            cell = ws.cell(row=r, column=4)
            # Check format
            if not is_dd_mmm_yyyy_exact(cell.number_format):
                format_fail_count += 1
            # Check value
            expected = EXPECTED_DATES[i]
            if not isinstance(cell.value, datetime.datetime):
                value_fail_count += 1
                mismatches.append(f"D{r}: expected datetime, got {type(cell.value).__name__}={cell.value!r}")
            elif cell.value.date() != expected.date():
                value_fail_count += 1
                mismatches.append(f"D{r}: expected {expected.date()}, got {cell.value.date()}")

        if format_fail_count == 0 and value_fail_count == 0:
            print(f"PASS: Component 3 — DD-MMM-YYYY format applied AND all date values preserved (0.2 pts)")
            total_score += 0.2
        elif format_fail_count > 0:
            print(f"FAIL: Component 3 — format is not DD-MMM-YYYY on {format_fail_count} cells")
        else:
            print(f"FAIL: Component 3 — date values not preserved in {value_fail_count} cells")
            for m in mismatches[:3]:
                print(f"  {m}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
