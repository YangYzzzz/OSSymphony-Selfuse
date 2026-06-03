"""
Reward Script: Create a pivot table from attendance data
Task ID: calc_pivot_019
Domain: libreoffice_calc
Scoring:
  Component 1: Pivot table sheet exists (0.2 pts)
  Component 2: Correct structure — 10 employee rows, 6 month columns (0.3 pts)
  Component 3: Correct attendance count values (0.3 pts)
  Component 4: Grand Total row with correct totals (0.2 pts)
"""

import os
import openpyxl
from collections import Counter
from datetime import datetime

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_019'


def compute_expected_pivot(wb):
    """Compute expected pivot from AttendanceLog data."""
    ws = wb['AttendanceLog']
    month_names = {1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June'}

    # Build pivot: {employee: {month_name: count}}
    pivot = {}
    for r in range(2, ws.max_row + 1):
        emp = ws.cell(r, 2).value
        date_val = ws.cell(r, 3).value
        if emp and date_val:
            if isinstance(date_val, datetime):
                month = date_val.month
            else:
                continue
            month_name = month_names.get(month, str(month))
            if emp not in pivot:
                pivot[emp] = Counter()
            pivot[emp][month_name] += 1

    return pivot, month_names


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: AttendanceLog must still exist
    if 'AttendanceLog' not in wb.sheetnames:
        print("FAIL: AttendanceLog sheet missing — file may be corrupted")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: A pivot table sheet exists (0.2 points)
    # The golden has a sheet named "PivotTable" but we accept any new sheet
    # that is not "AttendanceLog" and contains tabular data resembling a pivot.
    try:
        pivot_sheet = None
        pivot_sheet_name = None
        for sn in wb.sheetnames:
            if sn != 'AttendanceLog':
                candidate = wb[sn]
                # Check it has at least some data (more than just a header)
                if candidate.max_row >= 3 and candidate.max_column >= 3:
                    pivot_sheet = candidate
                    pivot_sheet_name = sn
                    break

        if pivot_sheet is not None:
            print(f"PASS: Component 1 — Pivot table sheet '{pivot_sheet_name}' exists (0.2 pts)")
            total_score += 0.2
        else:
            print("FAIL: Component 1 — No pivot table sheet found (need a new sheet with tabular data)")
            print(f"  Sheets found: {wb.sheetnames}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    if pivot_sheet is None:
        # Cannot continue without a pivot sheet
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    # Compute expected pivot from source data
    try:
        expected_pivot, month_names = compute_expected_pivot(wb)
    except Exception as e:
        print(f"ERROR: Could not compute expected pivot: {e}")
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    expected_employees = sorted(expected_pivot.keys())
    expected_months = ['January', 'February', 'March', 'April', 'May', 'June']

    # Component 2: Correct structure — employee rows and month columns (0.3 points)
    try:
        ws = pivot_sheet

        # Read header row to find month columns
        header_row = []
        for c in range(1, ws.max_column + 1):
            val = ws.cell(1, c).value
            if val is not None:
                header_row.append(str(val).strip())
            else:
                header_row.append('')

        # Find month columns (case-insensitive partial match)
        month_col_map = {}  # month_name -> column index (1-based)
        for c_idx, h in enumerate(header_row):
            for m in expected_months:
                if m.lower() in h.lower() or h.lower() in m.lower():
                    month_col_map[m] = c_idx + 1  # 1-based
                    break

        # Find employee name column (first column or column labeled with employee-like header)
        emp_col = 1  # default

        # Read employee names from pivot sheet
        found_employees = []
        for r in range(2, ws.max_row + 1):
            val = ws.cell(r, emp_col).value
            if val is not None and str(val).strip().lower() not in ('grand total', 'total', ''):
                found_employees.append(str(val).strip())

        # Check structure
        months_found = len(month_col_map)
        employees_found = len(found_employees)

        structure_score = 0.0
        # Sub-check: months present (0.15 pts)
        if months_found >= 6:
            structure_score += 0.15
            print(f"PASS: Component 2a — All 6 months found as columns ({list(month_col_map.keys())})")
        elif months_found >= 3:
            structure_score += 0.075
            print(f"PARTIAL: Component 2a — {months_found}/6 months found")
        else:
            print(f"FAIL: Component 2a — Only {months_found}/6 months found in header: {header_row}")

        # Sub-check: employees present (0.15 pts)
        matched_employees = set()
        for fe in found_employees:
            for ee in expected_employees:
                if fe.lower() == ee.lower():
                    matched_employees.add(ee)

        if len(matched_employees) >= 10:
            structure_score += 0.15
            print(f"PASS: Component 2b — All 10 employees found")
        elif len(matched_employees) >= 5:
            structure_score += 0.075
            print(f"PARTIAL: Component 2b — {len(matched_employees)}/10 employees found")
        else:
            print(f"FAIL: Component 2b — Only {len(matched_employees)}/10 employees found")
            print(f"  Found: {found_employees[:5]}")
            print(f"  Expected: {expected_employees[:5]}")

        total_score += structure_score
        print(f"  Component 2 total: {structure_score}/0.3 pts")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Correct attendance count values (0.3 points)
    try:
        ws = pivot_sheet
        correct_cells = 0
        total_cells = 0

        # Build a mapping of pivot sheet employee rows
        emp_row_map = {}  # employee_name -> row index
        for r in range(2, ws.max_row + 1):
            val = ws.cell(r, emp_col).value
            if val is not None and str(val).strip().lower() not in ('grand total', 'total', ''):
                emp_row_map[str(val).strip()] = r

        for emp_name in expected_employees:
            # Find this employee in the pivot sheet
            pivot_row = None
            for pe, pr in emp_row_map.items():
                if pe.lower() == emp_name.lower():
                    pivot_row = pr
                    break

            if pivot_row is None:
                total_cells += len(expected_months)
                continue

            for month_name in expected_months:
                if month_name not in month_col_map:
                    total_cells += 1
                    continue

                col_idx = month_col_map[month_name]
                cell_val = ws.cell(pivot_row, col_idx).value
                expected_val = expected_pivot[emp_name].get(month_name, 0)
                total_cells += 1

                if cell_val is not None:
                    try:
                        if abs(float(cell_val) - expected_val) < 0.5:
                            correct_cells += 1
                    except (ValueError, TypeError):
                        pass

        if total_cells > 0:
            accuracy = correct_cells / total_cells
            value_score = round(0.3 * accuracy, 4)
            total_score += value_score
            print(f"{'PASS' if accuracy >= 0.95 else 'PARTIAL' if accuracy > 0 else 'FAIL'}: Component 3 — {correct_cells}/{total_cells} cells correct (accuracy: {accuracy:.1%}) ({value_score} pts)")
        else:
            print("FAIL: Component 3 — No cells to verify")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Grand Total row with correct totals (0.2 points)
    try:
        ws = pivot_sheet
        grand_total_row = None

        # Find grand total row
        for r in range(2, ws.max_row + 1):
            val = ws.cell(r, emp_col).value
            if val is not None and str(val).strip().lower() in ('grand total', 'total'):
                grand_total_row = r
                break

        if grand_total_row is None:
            print("FAIL: Component 4 — No Grand Total row found")
        else:
            gt_score = 0.0

            # Check that Grand Total row exists (0.1 pts)
            gt_score += 0.1

            # Check grand total value (rightmost or sum column) = 600 (0.1 pts)
            # Check the last column (Grand Total column)
            last_col = ws.max_column
            gt_last_val = ws.cell(grand_total_row, last_col).value

            # Also check month totals
            month_totals_correct = 0
            month_totals_checked = 0
            for month_name in expected_months:
                if month_name not in month_col_map:
                    continue
                col_idx = month_col_map[month_name]
                cell_val = ws.cell(grand_total_row, col_idx).value
                expected_month_total = sum(expected_pivot[emp].get(month_name, 0) for emp in expected_employees)
                month_totals_checked += 1
                if cell_val is not None:
                    try:
                        if abs(float(cell_val) - expected_month_total) < 0.5:
                            month_totals_correct += 1
                    except (ValueError, TypeError):
                        pass

            # Grand total of 600
            gt_val_matches = (
                gt_last_val is not None
                and abs(float(gt_last_val) - 600) < 0.5
            ) if gt_last_val is not None else False

            if gt_val_matches and month_totals_checked > 0 and month_totals_correct == month_totals_checked:
                gt_score += 0.1
                print(f"PASS: Component 4 — Grand Total row correct (overall: {gt_last_val}, month totals: {month_totals_correct}/{month_totals_checked}) (0.2 pts)")
            elif gt_val_matches:
                gt_score += 0.05
                print(f"PARTIAL: Component 4 — Grand Total overall correct ({gt_last_val}) but month totals: {month_totals_correct}/{month_totals_checked} ({gt_score} pts)")
            else:
                print(f"FAIL: Component 4 — Grand Total value incorrect. Found: {gt_last_val}, expected: 600")

            total_score += gt_score
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook for LibreOffice
def persist_app_state(domain):
    import os, time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


# Main execution
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
