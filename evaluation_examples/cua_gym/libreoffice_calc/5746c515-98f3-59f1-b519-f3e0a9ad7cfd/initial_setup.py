"""
Initial Setup: Create attendance log spreadsheet for pivot table task
Task ID: calc_pivot_019
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_019'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'AttendanceLog'

    # --- Headers ---
    headers = ['RecordID', 'EmployeeName', 'Date', 'Status']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Employee list (10 employees) ---
    employees = [
        'Sarah Chen',
        'Marcus Johnson',
        'Priya Patel',
        'David Kim',
        'Elena Rodriguez',
        'James O\'Brien',
        'Aisha Mohammed',
        'Lucas Weber',
        'Mei-Lin Huang',
        'Carlos Fernandez',
    ]

    statuses = ['Present', 'Absent', 'Late', 'WFH']
    status_weights = [60, 10, 15, 15]  # weighted distribution

    # --- Generate 600 rows of attendance data ---
    # Jan-Jun 2024, ~100 working days, 10 employees = ~1000 possible entries
    # We pick ~60 entries per employee (600 total) spread across Jan-Jun

    # Build all weekdays Jan-Jun 2024
    start_date = date(2024, 1, 1)
    end_date = date(2024, 6, 30)
    all_days = []
    current = start_date
    while current <= end_date:
        if current.weekday() < 5:  # Mon-Fri
            all_days.append(current)
        current += timedelta(days=1)

    # For each employee, pick ~60 days from all_days
    records = []
    record_id = 1
    for emp in employees:
        # Sample 60 days per employee to get 600 total
        days_for_emp = sorted(random.sample(all_days, 60))
        for d in days_for_emp:
            status = random.choices(statuses, weights=status_weights, k=1)[0]
            records.append((record_id, emp, d, status))
            record_id += 1

    # Shuffle to make it look like a real log (not grouped by employee)
    random.shuffle(records)

    # Re-assign sequential RecordIDs after shuffle
    for idx, (_, emp, d, status) in enumerate(records):
        row = idx + 2
        ws.cell(row=row, column=1, value=idx + 1)
        ws.cell(row=row, column=2, value=emp)
        ws.cell(row=row, column=3, value=d)
        ws.cell(row=row, column=3).number_format = 'yyyy-mm-dd'
        ws.cell(row=row, column=4, value=status)

    # --- Column widths ---
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12

    # --- Freeze header row ---
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
