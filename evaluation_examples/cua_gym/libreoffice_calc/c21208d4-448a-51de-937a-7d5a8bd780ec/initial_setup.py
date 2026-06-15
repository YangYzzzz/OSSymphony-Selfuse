"""
Initial Setup: Create Support_Tickets spreadsheet with ticket data (no conditional formatting)
Task ID: calc_gcv_016
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_016'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Support_Tickets"

    # Headers
    ws.cell(row=1, column=1, value="Ticket Number")
    ws.cell(row=1, column=2, value="Subject")

    # Column widths for readability
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 50

    # 39 rows of ticket data (rows 2-40)
    tickets = [
        ("TK-1001", "URGENT: Server down in production cluster"),
        ("TK-1002", "Re: Account access request for new hire"),
        ("TK-1003", "URGENT: Data loss reported in finance database"),
        ("TK-1004", "Weekly backup schedule confirmation"),
        ("TK-1005", "URGENT: Email service outage affecting all departments"),
        ("TK-1006", "Request for additional monitor setup"),
        ("TK-1007", "VPN connection issue from remote office"),
        ("TK-1008", "URGENT: Security breach detected on web portal"),
        ("TK-1009", "Software license renewal for Adobe Creative Suite"),
        ("TK-1010", "Printer jam on 3rd floor - HP LaserJet Pro"),
        ("TK-1011", "URGENT: Customer payment gateway failure"),
        ("TK-1012", "New employee onboarding IT checklist"),
        ("TK-1013", "Request to upgrade RAM on workstation WS-2847"),
        ("TK-1014", "URGENT: Database replication lag exceeding threshold"),
        ("TK-1015", "Conference room AV system not responding"),
        ("TK-1016", "Slack integration with Jira broken since update"),
        ("TK-1017", "URGENT: SSL certificate expiring in 24 hours"),
        ("TK-1018", "Shared drive permissions for marketing team"),
        ("TK-1019", "Office 365 sync error on multiple devices"),
        ("TK-1020", "URGENT: Firewall rule misconfiguration blocking API"),
        ("TK-1021", "Re: Follow up on laptop replacement program"),
        ("TK-1022", "Network speed degradation in Building B"),
        ("TK-1023", "URGENT: Production deployment rollback needed"),
        ("TK-1024", "Request for USB-C docking station"),
        ("TK-1025", "URGENT: Client-facing dashboard showing stale data"),
        ("TK-1026", "Badge access not working for new employees"),
        ("TK-1027", "Scheduled maintenance window approval for Saturday"),
        ("TK-1028", "URGENT: Memory leak in order processing service"),
        ("TK-1029", "Wi-Fi connectivity drops in cafeteria area"),
        ("TK-1030", "Re: Software installation request - Tableau Desktop"),
        ("TK-1031", "URGENT: Backup job failed for three consecutive nights"),
        ("TK-1032", "Desk phone not receiving calls - Extension 4521"),
        ("TK-1033", "Request for dual-boot setup on development machine"),
        ("TK-1034", "URGENT: DNS resolution failures across network"),
        ("TK-1035", "New project shared folder creation request"),
        ("TK-1036", "Screen flickering on Dell monitor - Dock issue"),
        ("TK-1037", "URGENT: Critical patch pending for CVE-2025-3891"),
        ("TK-1038", "Inventory system showing incorrect stock counts"),
        ("TK-1039", "Meeting room booking system calendar sync issue"),
    ]

    for i, (ticket_num, subject) in enumerate(tickets, 2):
        ws.cell(row=i, column=1, value=ticket_num)
        ws.cell(row=i, column=2, value=subject)

    # No conditional formatting in initial state
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
