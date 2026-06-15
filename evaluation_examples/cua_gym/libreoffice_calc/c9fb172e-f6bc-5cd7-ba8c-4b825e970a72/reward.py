"""
Reward Script: Protect document structure with password 'struct789'
Task ID: calc_ps_004
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): lockStructure is True
  Component 2 (0.3): workbookPassword is set (non-empty hash)
  Component 3 (0.3): Individual sheets NOT protected (contents remain editable)
                      combined with lockStructure being True (anchored to task change)
"""

import os

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_004'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice document."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: workbook must have the expected sheets
    expected_sheets = ['Summary', 'Details', 'Archive']
    if sorted(wb.sheetnames) != sorted(expected_sheets):
        print(f"WARN: Unexpected sheets. Expected {expected_sheets}, found {wb.sheetnames}")
        # Don't gate on this — sheets may have been renamed but protection still applied

    security = wb.security

    # Component 1: lockStructure is True (0.4 points)
    # This is the core requirement — structure protection prevents insert/delete/rename/reorder
    # INITIAL: lockStructure is None  -> FAIL
    # GOLDEN:  lockStructure is True  -> PASS
    try:
        lock_struct = security.lockStructure
        if lock_struct is True:
            print(f"PASS: Component 1 — lockStructure is True (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — lockStructure expected True, found: {lock_struct}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: workbookPassword is set (0.3 points)
    # The task specifies password 'struct789'. openpyxl stores a hash, not plaintext.
    # We verify a non-empty password hash exists (password was set).
    # INITIAL: workbookPassword is None -> FAIL
    # GOLDEN:  workbookPassword is '8B54' (or similar hash) -> PASS
    try:
        wb_password = security.workbookPassword
        if wb_password is not None and str(wb_password).strip() != '':
            print(f"PASS: Component 2 — workbookPassword is set (hash: {wb_password}) (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — workbookPassword expected non-empty, found: {wb_password}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Structure is protected BUT individual sheets are NOT protected (0.3 points)
    # Task says "Sheet contents remain editable unless individually protected"
    # This compound check: lockStructure must be True AND no individual sheet has sheet protection enabled
    # INITIAL: lockStructure is None -> compound fails -> FAIL
    # GOLDEN:  lockStructure is True AND sheets not individually protected -> PASS
    try:
        if security.lockStructure is True:
            individually_protected = [
                name for name in wb.sheetnames
                if wb[name].protection.sheet is True
            ]
            for name in individually_protected:
                print(f"  INFO: Sheet '{name}' has individual protection enabled")
            if len(individually_protected) == 0:
                print(f"PASS: Component 3 — Structure protected, sheet contents remain editable (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 3 — {len(individually_protected)} sheet(s) have individual protection (should remain editable)")
        else:
            print(f"FAIL: Component 3 — lockStructure is not True, compound check fails")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
