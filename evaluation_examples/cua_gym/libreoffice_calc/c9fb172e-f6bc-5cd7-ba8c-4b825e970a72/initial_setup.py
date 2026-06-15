"""
Initial Setup: Create workbook with three sheets (Summary, Details, Archive)
without structure protection.
Task ID: calc_ps_004
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_004'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Styling helpers ---
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    def write_headers(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def write_data(ws, data, start_row=2):
        for r, row_data in enumerate(data, start_row):
            for c, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = thin_border

    # ===== Sheet 1: Summary =====
    ws1 = wb.active
    ws1.title = 'Summary'

    summary_headers = ['Department', 'Q1 Revenue', 'Q2 Revenue', 'Q3 Revenue', 'Q4 Revenue', 'Annual Total']
    write_headers(ws1, summary_headers)

    summary_data = [
        ['Engineering', 245000, 267500, 289000, 312000, 1113500],
        ['Marketing', 180000, 195000, 210000, 225000, 810000],
        ['Sales', 320000, 345000, 375000, 410000, 1450000],
        ['Human Resources', 95000, 98000, 101000, 105000, 399000],
        ['Finance', 150000, 155000, 162000, 170000, 637000],
        ['Operations', 210000, 220000, 235000, 248000, 913000],
        ['Legal', 120000, 125000, 130000, 135000, 510000],
        ['Customer Support', 175000, 182000, 190000, 198000, 745000],
        ['Research & Development', 280000, 295000, 310000, 330000, 1215000],
        ['Product Management', 165000, 172000, 180000, 188000, 705000],
    ]
    write_data(ws1, summary_data)

    # Column widths
    ws1.column_dimensions['A'].width = 24
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws1.column_dimensions[col_letter].width = 16

    # Number format for currency columns
    for r in range(2, 12):
        for c in range(2, 7):
            ws1.cell(row=r, column=c).number_format = '$#,##0'

    # ===== Sheet 2: Details =====
    ws2 = wb.create_sheet('Details')

    details_headers = ['Employee ID', 'Name', 'Department', 'Position', 'Hire Date', 'Base Salary', 'Bonus']
    write_headers(ws2, details_headers)

    details_data = [
        ['EMP-1001', 'Sarah Chen', 'Engineering', 'Senior Developer', '2022-03-15', 95000, 12000],
        ['EMP-1002', 'Marcus Johnson', 'Marketing', 'Campaign Manager', '2021-07-01', 78000, 8500],
        ['EMP-1003', 'Priya Patel', 'Sales', 'Regional Director', '2020-01-20', 105000, 22000],
        ['EMP-1004', 'David Kim', 'Engineering', 'Tech Lead', '2019-11-10', 115000, 15000],
        ['EMP-1005', 'Laura Martinez', 'Human Resources', 'HR Manager', '2021-04-05', 82000, 7000],
        ['EMP-1006', 'James Wilson', 'Finance', 'Senior Analyst', '2022-08-22', 88000, 9500],
        ['EMP-1007', 'Aisha Rahman', 'Operations', 'Logistics Lead', '2020-06-15', 76000, 6800],
        ['EMP-1008', 'Robert Taylor', 'Legal', 'General Counsel', '2018-02-28', 135000, 18000],
        ['EMP-1009', 'Emily Zhang', 'Customer Support', 'Support Director', '2021-09-12', 85000, 8000],
        ['EMP-1010', 'Carlos Rivera', 'R&D', 'Research Scientist', '2023-01-08', 98000, 11000],
        ['EMP-1011', 'Hannah Brooks', 'Product Management', 'Product Owner', '2022-05-20', 92000, 10500],
        ['EMP-1012', 'Tomoko Sato', 'Engineering', 'DevOps Engineer', '2023-04-14', 90000, 9000],
    ]
    write_data(ws2, details_data)

    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 22
    ws2.column_dimensions['D'].width = 20
    ws2.column_dimensions['E'].width = 14
    ws2.column_dimensions['F'].width = 14
    ws2.column_dimensions['G'].width = 12

    for r in range(2, 14):
        ws2.cell(row=r, column=6).number_format = '$#,##0'
        ws2.cell(row=r, column=7).number_format = '$#,##0'

    # ===== Sheet 3: Archive =====
    ws3 = wb.create_sheet('Archive')

    archive_headers = ['Record ID', 'Date', 'Description', 'Amount', 'Status', 'Approved By']
    write_headers(ws3, archive_headers)

    archive_data = [
        ['ARC-2024-001', '2024-01-15', 'Office equipment purchase', 12500, 'Completed', 'M. Johnson'],
        ['ARC-2024-002', '2024-02-03', 'Software license renewal', 45000, 'Completed', 'S. Chen'],
        ['ARC-2024-003', '2024-02-28', 'Q1 marketing campaign', 28000, 'Completed', 'P. Patel'],
        ['ARC-2024-004', '2024-03-10', 'Server infrastructure upgrade', 67000, 'Completed', 'D. Kim'],
        ['ARC-2024-005', '2024-04-05', 'Employee training program', 15800, 'Completed', 'L. Martinez'],
        ['ARC-2024-006', '2024-04-22', 'Annual audit fees', 32000, 'Completed', 'J. Wilson'],
        ['ARC-2024-007', '2024-05-18', 'Trade show participation', 21500, 'Completed', 'R. Taylor'],
        ['ARC-2024-008', '2024-06-01', 'Insurance premium payment', 48000, 'Completed', 'E. Zhang'],
        ['ARC-2024-009', '2024-07-12', 'Warehouse lease deposit', 36000, 'Completed', 'A. Rahman'],
        ['ARC-2024-010', '2024-08-30', 'Patent filing expenses', 8900, 'Completed', 'C. Rivera'],
    ]
    write_data(ws3, archive_data)

    ws3.column_dimensions['A'].width = 18
    ws3.column_dimensions['B'].width = 14
    ws3.column_dimensions['C'].width = 32
    ws3.column_dimensions['D'].width = 14
    ws3.column_dimensions['E'].width = 14
    ws3.column_dimensions['F'].width = 16

    for r in range(2, 12):
        ws3.cell(row=r, column=4).number_format = '$#,##0'

    # NO structure protection on initial
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
