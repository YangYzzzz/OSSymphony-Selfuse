"""
Reward Script: Protect 'Attendance' sheet with password, allowing insert/delete rows
Task ID: calc_ps_024
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Sheet protection is enabled (sheet=True)
  Component 2 (0.2): Password is set (non-empty)
  Component 3 (0.2): Insert rows is allowed (insertRows=False in protection)
  Component 4 (0.2): Delete rows is allowed (deleteRows=False in protection)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_024'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Attendance' sheet must exist
    if 'Attendance' not in wb.sheetnames:
        print(f"FAIL: 'Attendance' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Attendance']
    prot = ws.protection

    # Component 1: Sheet protection is enabled (0.4 points)
    # Initial: sheet=False, Golden: sheet=True
    try:
        if prot.sheet is True:
            print(f"PASS: Component 1 — Sheet protection is enabled (sheet={prot.sheet}) (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — Sheet protection not enabled (sheet={prot.sheet})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Password is set (0.2 points)
    # Initial: password=None, Golden: password is a hash string
    try:
        has_password = (prot.password is not None and str(prot.password).strip() != '')
        if has_password:
            print(f"PASS: Component 2 — Password is set (hash={prot.password}) (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 2 — No password set (password={prot.password})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Insert rows is allowed (0.2 points)
    # In openpyxl SheetProtection, insertRows=False means inserting rows is ALLOWED
    # Initial: insertRows=True (blocked by default flag, but protection off so irrelevant)
    # Golden: insertRows=False (allowed while protection is on)
    # We need protection enabled AND insertRows=False
    try:
        if prot.sheet is True and prot.insertRows is False:
            print(f"PASS: Component 3 — Insert rows allowed (insertRows={prot.insertRows}) (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — Insert rows not properly configured (sheet={prot.sheet}, insertRows={prot.insertRows})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Delete rows is allowed (0.2 points)
    # Same logic: deleteRows=False means deleting rows is ALLOWED when protection is on
    # Initial: deleteRows=True (default), Golden: deleteRows=False (allowed)
    try:
        if prot.sheet is True and prot.deleteRows is False:
            print(f"PASS: Component 4 — Delete rows allowed (deleteRows={prot.deleteRows}) (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 4 — Delete rows not properly configured (sheet={prot.sheet}, deleteRows={prot.deleteRows})")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
