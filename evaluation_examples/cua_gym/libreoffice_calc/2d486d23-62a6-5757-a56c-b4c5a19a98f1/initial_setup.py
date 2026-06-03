"""
Initial Setup: Create attendance spreadsheet (unprotected)
Task ID: calc_ps_024
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_024'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attendance'

    # --- Headers ---
    headers = ['Employee', 'Date', 'Status', 'Hours']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2E75B6', end_color='FF2E75B6', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Column widths ---
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10

    # --- 40 rows of realistic attendance data ---
    employees = [
        'Sarah Chen', 'Marcus Johnson', 'Priya Patel', 'James Wilson',
        'Maria Garcia', 'David Kim', 'Emily Brown', 'Robert Taylor',
        'Aisha Mohammed', 'Thomas Anderson', 'Lisa Nakamura', 'Carlos Rivera',
        'Hannah Mueller', 'Wei Zhang', 'Olivia Foster', 'Daniel Okonkwo',
        'Sophie Martin', 'Ryan O\'Brien', 'Fatima Al-Hassan', 'Michael Scott',
    ]

    statuses = ['Present', 'Present', 'Present', 'Present', 'Present',
                'Present', 'Late', 'Late', 'Absent', 'Half Day']

    random.seed(42)
    base_date = date(2025, 3, 3)  # Monday

    data_align = Alignment(horizontal='center', vertical='center')

    for r in range(40):
        row_num = r + 2
        emp = employees[r % len(employees)]
        # Spread across 2 weeks of dates
        d = base_date + timedelta(days=(r // 20) * 7 + (r % 5))
        status = random.choice(statuses)
        if status == 'Present':
            hours = round(random.uniform(7.5, 9.0), 1)
        elif status == 'Late':
            hours = round(random.uniform(5.0, 7.5), 1)
        elif status == 'Half Day':
            hours = 4.0
        else:
            hours = 0.0

        ws.cell(row=row_num, column=1, value=emp).border = thin_border
        date_cell = ws.cell(row=row_num, column=2, value=d)
        date_cell.number_format = 'yyyy-mm-dd'
        date_cell.alignment = data_align
        date_cell.border = thin_border
        status_cell = ws.cell(row=row_num, column=3, value=status)
        status_cell.alignment = data_align
        status_cell.border = thin_border
        hours_cell = ws.cell(row=row_num, column=4, value=hours)
        hours_cell.number_format = '0.0'
        hours_cell.alignment = data_align
        hours_cell.border = thin_border

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Sheet is NOT protected (initial state)
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
