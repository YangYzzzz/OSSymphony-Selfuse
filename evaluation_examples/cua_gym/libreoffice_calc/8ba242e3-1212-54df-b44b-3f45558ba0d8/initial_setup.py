"""
Initial Setup: Inventory multi-source weighted average valuation
Task ID: calc_ops_inventory_multisource_valuation_043
Domain: libreoffice_calc

Creates:
  - Sheet 'ReceiptHistory': 150 receipt records for 40 SKUs from multiple suppliers
  - Sheet 'InventoryMaster': 40 SKUs with empty Weighted Avg Cost column (D)
"""

import os
import random
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_inventory_multisource_valuation_043'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

random.seed(42)

# --- 40 unique SKUs with product info ---
PRODUCTS = [
    ('SKU-1001', 'Industrial Bearing Set 6205', 'Mechanical Parts'),
    ('SKU-1002', 'Stainless Steel Bolt M10x50', 'Fasteners'),
    ('SKU-1003', 'Hydraulic Seal Kit 45mm', 'Seals & Gaskets'),
    ('SKU-1004', 'Copper Pipe Fitting 3/4"', 'Plumbing'),
    ('SKU-1005', 'Safety Valve 1000 PSI', 'Pressure Control'),
    ('SKU-1006', 'Electric Motor 2HP 3-Phase', 'Electrical'),
    ('SKU-1007', 'V-Belt B68 Industrial', 'Drive Components'),
    ('SKU-1008', 'Roller Chain #50 10ft', 'Drive Components'),
    ('SKU-1009', 'Air Filter Element AF2550', 'Filtration'),
    ('SKU-1010', 'Oil Seal 25x40x7mm', 'Seals & Gaskets'),
    ('SKU-1011', 'Hex Nut M12 Grade 8', 'Fasteners'),
    ('SKU-1012', 'Pneumatic Cylinder 80mm Bore', 'Pneumatics'),
    ('SKU-1013', 'Flow Control Valve 1/2"', 'Fluid Control'),
    ('SKU-1014', 'Proximity Sensor PNP 10-30V', 'Sensors'),
    ('SKU-1015', 'Circuit Breaker 20A 3-Pole', 'Electrical'),
    ('SKU-1016', 'Lubricant Grease NLGI-2 1kg', 'Lubricants'),
    ('SKU-1017', 'Timing Belt HTD 5M 15mm', 'Drive Components'),
    ('SKU-1018', 'Gasket Sheet 500x500 2mm', 'Seals & Gaskets'),
    ('SKU-1019', 'Pressure Gauge 0-10 Bar 63mm', 'Instrumentation'),
    ('SKU-1020', 'Terminal Block DIN Rail 10A', 'Electrical'),
    ('SKU-1021', 'Conveyor Belt PVC 600mm', 'Material Handling'),
    ('SKU-1022', 'Solenoid Valve 24VDC 1/4"', 'Pneumatics'),
    ('SKU-1023', 'Angular Contact Bearing 7205', 'Mechanical Parts'),
    ('SKU-1024', 'Stainless Washer M8 Grade A4', 'Fasteners'),
    ('SKU-1025', 'Cooling Fan 80x80x25mm 24V', 'Cooling'),
    ('SKU-1026', 'Level Sensor Ultrasonic 5m', 'Sensors'),
    ('SKU-1027', 'PLC Module Digital I/O 16pt', 'Automation'),
    ('SKU-1028', 'Flexible Conduit 20mm x 25m', 'Electrical'),
    ('SKU-1029', 'Rotary Encoder 1000 PPR', 'Sensors'),
    ('SKU-1030', 'Sight Glass 1" NPT Borosilicate', 'Instrumentation'),
    ('SKU-1031', 'Gear Pump 10 LPM', 'Fluid Control'),
    ('SKU-1032', 'Spring Lock Washer M6 Zinc', 'Fasteners'),
    ('SKU-1033', 'PTFE Tape 12mm x 10m', 'Sealing Materials'),
    ('SKU-1034', 'Flexible Coupling 14mm Bore', 'Drive Components'),
    ('SKU-1035', 'Contactor 18A 24VDC Coil', 'Electrical'),
    ('SKU-1036', 'Heat Exchanger Plate Type 10kW', 'Thermal'),
    ('SKU-1037', 'Cable Tray 100x50mm 3m', 'Electrical'),
    ('SKU-1038', 'Diaphragm Pump 40 LPM', 'Fluid Control'),
    ('SKU-1039', 'Limit Switch Roller Lever', 'Sensors'),
    ('SKU-1040', 'Vibration Damper 25x25mm M8', 'Mechanical Parts'),
]

SUPPLIERS = [
    'Apex Industrial Supply',
    'Pacific Parts & Components',
    'TechMech Distributors',
    'Global Machinery Parts',
    'Premier Industrial Co.',
    'FastTrack Supplies Ltd',
    'Continental Parts Group',
    'Alliance Technical Supply',
]

# Base prices per SKU (realistic unit costs in USD)
BASE_PRICES = {
    'SKU-1001': 18.50, 'SKU-1002': 0.85, 'SKU-1003': 32.40, 'SKU-1004': 4.25,
    'SKU-1005': 145.00, 'SKU-1006': 385.00, 'SKU-1007': 12.75, 'SKU-1008': 28.90,
    'SKU-1009': 22.60, 'SKU-1010': 5.40, 'SKU-1011': 0.35, 'SKU-1012': 89.50,
    'SKU-1013': 67.30, 'SKU-1014': 54.80, 'SKU-1015': 38.20, 'SKU-1016': 24.90,
    'SKU-1017': 31.50, 'SKU-1018': 16.75, 'SKU-1019': 42.00, 'SKU-1020': 3.85,
    'SKU-1021': 215.00, 'SKU-1022': 72.40, 'SKU-1023': 28.60, 'SKU-1024': 0.55,
    'SKU-1025': 19.80, 'SKU-1026': 183.50, 'SKU-1027': 247.00, 'SKU-1028': 44.20,
    'SKU-1029': 138.90, 'SKU-1030': 56.70, 'SKU-1031': 320.00, 'SKU-1032': 0.18,
    'SKU-1033': 2.95, 'SKU-1034': 28.40, 'SKU-1035': 48.60, 'SKU-1036': 890.00,
    'SKU-1037': 35.60, 'SKU-1038': 468.00, 'SKU-1039': 22.30, 'SKU-1040': 8.75,
}

# Current stock quantities per SKU
STOCK = {
    'SKU-1001': 245, 'SKU-1002': 1820, 'SKU-1003': 58, 'SKU-1004': 340,
    'SKU-1005': 15, 'SKU-1006': 8, 'SKU-1007': 92, 'SKU-1008': 47,
    'SKU-1009': 124, 'SKU-1010': 380, 'SKU-1011': 3240, 'SKU-1012': 22,
    'SKU-1013': 33, 'SKU-1014': 61, 'SKU-1015': 45, 'SKU-1016': 112,
    'SKU-1017': 76, 'SKU-1018': 88, 'SKU-1019': 54, 'SKU-1020': 520,
    'SKU-1021': 14, 'SKU-1022': 39, 'SKU-1023': 96, 'SKU-1024': 2100,
    'SKU-1025': 73, 'SKU-1026': 19, 'SKU-1027': 12, 'SKU-1028': 68,
    'SKU-1029': 24, 'SKU-1030': 41, 'SKU-1031': 11, 'SKU-1032': 5600,
    'SKU-1033': 430, 'SKU-1034': 55, 'SKU-1035': 36, 'SKU-1036': 4,
    'SKU-1037': 62, 'SKU-1038': 9, 'SKU-1039': 88, 'SKU-1040': 150,
}


def random_date(start_date, end_date):
    delta = end_date - start_date
    random_days = random.randint(0, delta.days)
    return start_date + timedelta(days=random_days)


def create_initial():
    wb = openpyxl.Workbook()

    # ---- Sheet 1: ReceiptHistory ----
    ws_rh = wb.active
    ws_rh.title = 'ReceiptHistory'

    # Headers
    rh_headers = ['SKU', 'Supplier', 'Qty Received', 'Unit Price Paid', 'Receipt Date']
    for col, h in enumerate(rh_headers, 1):
        cell = ws_rh.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Format columns
    ws_rh.column_dimensions['A'].width = 14
    ws_rh.column_dimensions['B'].width = 30
    ws_rh.column_dimensions['C'].width = 16
    ws_rh.column_dimensions['D'].width = 18
    ws_rh.column_dimensions['E'].width = 16

    # Generate 150 receipt records: assign roughly 3-5 receipts per SKU
    # (40 SKUs x ~3.75 avg = 150 records)
    start_date = date(2025, 1, 2)
    end_date = date(2025, 3, 31)

    receipt_rows = []
    # Ensure each SKU gets at least 2 receipts from different suppliers
    for sku, prod_name, cat in PRODUCTS:
        base = BASE_PRICES[sku]
        # 2 guaranteed receipts per SKU from different suppliers
        sup_pool = random.sample(SUPPLIERS, k=min(4, len(SUPPLIERS)))
        n_receipts = random.randint(2, 5)
        for i in range(n_receipts):
            supplier = sup_pool[i % len(sup_pool)]
            # Supplier variation: ±8% of base price
            price_variation = random.uniform(-0.08, 0.08)
            unit_price = round(base * (1 + price_variation), 4)
            qty = random.randint(10, 300)
            receipt_date = random_date(start_date, end_date)
            receipt_rows.append((sku, supplier, qty, unit_price, receipt_date.strftime('%Y-%m-%d')))

    # If we have more than 150, trim; if fewer, add extras
    random.shuffle(receipt_rows)
    # Trim or pad to exactly 150
    while len(receipt_rows) < 150:
        # Add extra receipts from random SKUs
        sku, prod_name, cat = random.choice(PRODUCTS)
        base = BASE_PRICES[sku]
        supplier = random.choice(SUPPLIERS)
        price_variation = random.uniform(-0.08, 0.08)
        unit_price = round(base * (1 + price_variation), 4)
        qty = random.randint(10, 200)
        receipt_date = random_date(start_date, end_date)
        receipt_rows.append((sku, supplier, qty, unit_price, receipt_date.strftime('%Y-%m-%d')))

    receipt_rows = receipt_rows[:150]

    for row_idx, (sku, supplier, qty, price, rec_date) in enumerate(receipt_rows, 2):
        ws_rh.cell(row=row_idx, column=1, value=sku)
        ws_rh.cell(row=row_idx, column=2, value=supplier)
        ws_rh.cell(row=row_idx, column=3, value=qty)
        cell_price = ws_rh.cell(row=row_idx, column=4, value=price)
        cell_price.number_format = '$#,##0.0000'
        ws_rh.cell(row=row_idx, column=5, value=rec_date)

    ws_rh.freeze_panes = 'A2'

    # ---- Sheet 2: InventoryMaster ----
    ws_im = wb.create_sheet('InventoryMaster')

    # Headers
    im_headers = ['SKU', 'Product Name', 'Category', 'Weighted Avg Cost', 'Current Stock']
    for col, h in enumerate(im_headers, 1):
        cell = ws_im.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    ws_im.column_dimensions['A'].width = 14
    ws_im.column_dimensions['B'].width = 38
    ws_im.column_dimensions['C'].width = 22
    ws_im.column_dimensions['D'].width = 20
    ws_im.column_dimensions['E'].width = 16

    # Data rows (D column = Weighted Avg Cost is EMPTY — task is to fill this)
    for row_idx, (sku, prod_name, cat) in enumerate(PRODUCTS, 2):
        ws_im.cell(row=row_idx, column=1, value=sku)
        ws_im.cell(row=row_idx, column=2, value=prod_name)
        ws_im.cell(row=row_idx, column=3, value=cat)
        # Column D (Weighted Avg Cost) intentionally LEFT EMPTY
        ws_im.cell(row=row_idx, column=5, value=STOCK[sku])

    ws_im.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  ReceiptHistory: 150 receipt records for 40 SKUs')
    print(f'  InventoryMaster: 40 SKUs, Weighted Avg Cost column (D) is empty')


create_initial()
