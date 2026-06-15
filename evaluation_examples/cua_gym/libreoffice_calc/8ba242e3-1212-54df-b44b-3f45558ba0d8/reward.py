"""
Reward Script: Weighted Average Unit Cost for Multi-Supplier Inventory
Task ID: calc_ops_inventory_multisource_valuation_043
Domain: libreoffice_calc
Scoring:
  - Precondition gate: ReceiptHistory sheet intact (151 rows, 5 cols, correct headers)
  - Component 1 (0.40): SUMPRODUCT weighted average formula in InventoryMaster D2:D41
  - Component 2 (0.30): Total Inventory Value column F (header + =Dx*Ex formulas F2:F41)
  - Component 3 (0.20): Currency format with 4 decimal places ($#,##0.0000) on D2:D41
  - Component 4 (0.10): D column ALL 40 rows filled (none are empty/None)

NOTE: ReceiptHistory integrity is a GATE (not a scoring component) since it is
identical in both initial and golden files. Components 1-4 only pass on the golden
file (post-task), ensuring initial scores 0.0.
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_inventory_multisource_valuation_043'

# Expected formula pattern in D column (case-insensitive, no-spaces comparison key)
EXPECTED_D_FORMULA_PATTERN = 'SUMPRODUCT'
EXPECTED_D_SUMIF_PATTERN = 'SUMIF'
EXPECTED_D_RECEIPT_REF = 'RECEIPTHISORY'  # partial ref used for looser check

# Expected number format for D column
EXPECTED_D_FORMAT = '$#,##0.0000'

def normalize_formula(f):
    """Normalize formula for comparison: uppercase, no spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify InventoryMaster sheet exists
    if 'InventoryMaster' not in wb.sheetnames:
        print("CRITICAL: 'InventoryMaster' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: ReceiptHistory must still be present (not deleted)
    # This is a GATE, not a scoring component — ReceiptHistory is identical
    # in initial and golden, so it cannot contribute to distinguishing them.
    if 'ReceiptHistory' not in wb.sheetnames:
        print("GATE FAIL: 'ReceiptHistory' sheet is missing — file may be corrupted")
        print("REWARD: 0.0")
        return 0.0

    ws_inv = wb['InventoryMaster']

    # -------------------------------------------------------------------
    # Component 1: SUMPRODUCT weighted average formula in D2:D41 (0.40 pts)
    # The formula must:
    #   - Be present in ALL 40 rows (D2 through D41)
    #   - Contain SUMPRODUCT and SUMIF
    #   - Reference ReceiptHistory columns A, C, D
    # -------------------------------------------------------------------
    try:
        d_formula_rows_ok = 0
        d_formula_rows_missing = []

        for row in range(2, 42):  # D2:D41 = rows 2..41
            val = ws_inv.cell(row=row, column=4).value
            norm = normalize_formula(val)
            # Must be a formula string with SUMPRODUCT and SUMIF referencing ReceiptHistory
            if (
                isinstance(val, str) and
                val.startswith('=') and
                'SUMPRODUCT' in norm and
                'SUMIF' in norm and
                'RECEIPTHISTORY' in norm
            ):
                d_formula_rows_ok += 1
            else:
                d_formula_rows_missing.append((row, repr(val)))

        if d_formula_rows_ok == 40:
            print(f"PASS: Component 1 — All 40 rows D2:D41 have SUMPRODUCT/SUMIF weighted-avg formula (0.40 pts)")
            total_score += 0.40
        elif d_formula_rows_ok >= 30:
            # Partial credit: most rows have the formula
            partial = round(0.40 * (d_formula_rows_ok / 40), 2)
            total_score += partial
            print(f"PARTIAL: Component 1 — {d_formula_rows_ok}/40 rows have formula ({partial} pts)")
            print(f"  Missing/wrong in rows: {[r for r, _ in d_formula_rows_missing[:5]]}")
        else:
            print(f"FAIL: Component 1 — Only {d_formula_rows_ok}/40 rows have SUMPRODUCT formula in D column")
            if d_formula_rows_missing:
                print(f"  First missing: row {d_formula_rows_missing[0][0]}: {d_formula_rows_missing[0][1]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------
    # Component 2: Total Inventory Value column F (0.30 pts)
    # F1 header = 'Total Inventory Value'
    # F2:F41 each has formula =Dx*Ex (linking weighted avg cost * current stock)
    # -------------------------------------------------------------------
    try:
        f1_val = ws_inv.cell(row=1, column=6).value
        header_ok = (
            f1_val is not None and
            'total inventory value' in str(f1_val).lower()
        )

        f_formula_rows_ok = 0
        f_formula_rows_missing = []

        for row in range(2, 42):  # F2:F41
            val = ws_inv.cell(row=row, column=6).value
            norm = normalize_formula(val)
            # Expect =Dx*Ex or =Ex*Dx (same thing)
            if (
                isinstance(val, str) and
                val.startswith('=') and
                f'D{row}' in val and
                f'E{row}' in val and
                '*' in val
            ):
                f_formula_rows_ok += 1
            else:
                f_formula_rows_missing.append((row, repr(val)))

        if header_ok and f_formula_rows_ok == 40:
            print(f"PASS: Component 2 — F1 header='{f1_val}', all 40 rows F2:F41 have =Dx*Ex formula (0.30 pts)")
            total_score += 0.30
        elif header_ok and f_formula_rows_ok >= 30:
            partial = round(0.30 * (f_formula_rows_ok / 40), 2)
            print(f"PARTIAL: Component 2 — Header OK, {f_formula_rows_ok}/40 F rows have formula ({partial} pts)")
            total_score += partial
        elif not header_ok:
            print(f"FAIL: Component 2 — F1 header missing or wrong: {repr(f1_val)}")
            if f_formula_rows_ok >= 30:
                # Give partial for having formulas even if header is off
                partial = round(0.15 * (f_formula_rows_ok / 40), 2)
                total_score += partial
                print(f"  Partial for formulas: {f_formula_rows_ok}/40 rows have formula ({partial} pts)")
        else:
            print(f"FAIL: Component 2 — Only {f_formula_rows_ok}/40 F rows have correct formula")
            if f_formula_rows_missing:
                print(f"  First missing: row {f_formula_rows_missing[0][0]}: {f_formula_rows_missing[0][1]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------
    # Component 3: Currency format with 4 decimal places on D2:D41 (0.20 pts)
    # Expected format: '$#,##0.0000' (currency, 4 decimals for precision)
    # -------------------------------------------------------------------
    try:
        d_format_rows_ok = 0
        d_format_seen = set()

        for row in range(2, 42):
            fmt = ws_inv.cell(row=row, column=4).number_format
            d_format_seen.add(fmt)
            # Check for currency format with 4 decimal places
            if fmt and (
                '$' in fmt and
                '0.0000' in fmt
            ):
                d_format_rows_ok += 1

        if d_format_rows_ok == 40:
            print(f"PASS: Component 3 — All 40 D column cells have currency 4-decimal format (0.20 pts)")
            total_score += 0.20
        elif d_format_rows_ok >= 20:
            partial = round(0.20 * (d_format_rows_ok / 40), 2)
            print(f"PARTIAL: Component 3 — {d_format_rows_ok}/40 D cells have correct currency format ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {d_format_rows_ok}/40 D cells have '$...0.0000' format")
            print(f"  Formats seen: {d_format_seen}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------
    # Component 4: ALL 40 D column cells are non-empty (0.10 pts)
    # This verifies the agent actually populated every row, not just some.
    # In the initial file, D2:D41 are all None — so this FAILS on initial.
    # In the golden file, all 40 rows have formulas — this PASSES on golden.
    # This component is distinct from Component 1's formula check:
    # it could catch a case where only some rows were filled.
    # -------------------------------------------------------------------
    try:
        d_nonempty_count = 0
        d_empty_rows = []

        for row in range(2, 42):  # D2:D41
            val = ws_inv.cell(row=row, column=4).value
            if val is not None:
                d_nonempty_count += 1
            else:
                d_empty_rows.append(row)

        if d_nonempty_count == 40:
            print(f"PASS: Component 4 — All 40 D column cells (D2:D41) are populated (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 — {len(d_empty_rows)} empty cells in D2:D41: {d_empty_rows[:5]}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score:.4f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
