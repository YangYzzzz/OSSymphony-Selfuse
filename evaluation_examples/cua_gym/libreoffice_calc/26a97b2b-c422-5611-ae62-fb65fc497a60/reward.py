"""
Reward Script: Protect the salary column so only HR managers with the password can edit it.
Task ID: calc_hr_041
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): Sheet protection is enabled
  Component 2 (0.30): C2:C15 cells are locked (salary data column)
  Component 3 (0.35): Non-salary cells (A,B,D columns + C1 header) are unlocked
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_041'


def persist_app_state(domain: str):
    """Save any unsaved edits in the GUI before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists
    if 'Confidential' not in wb.sheetnames:
        print("FAIL: 'Confidential' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Confidential']

    # Component 1: Sheet protection is enabled (0.35 points)
    # This FAILS on initial (protection.sheet == False) and PASSES on golden (True)
    try:
        if ws.protection.sheet:
            print(f"PASS: Component 1 — Sheet protection is enabled (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 1 — Sheet protection is NOT enabled (protection.sheet={ws.protection.sheet})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: C2:C15 cells are locked (0.30 points)
    # In initial env, all cells default to locked=True, but sheet protection is off so it doesn't matter.
    # In golden env, C2:C15 should be explicitly locked AND sheet protection is on.
    # We need this component to differentiate: it only awards points if BOTH sheet protection
    # is on AND C2:C15 are locked. This way initial env (protection off) scores 0.
    try:
        if ws.protection.sheet:
            locked_count = 0
            expected_locked = 14  # C2 through C15
            for row in range(2, 16):
                cell = ws.cell(row=row, column=3)  # column C
                if cell.protection.locked:
                    locked_count += 1
            if locked_count == expected_locked:
                print(f"PASS: Component 2 — All 14 salary cells (C2:C15) are locked ({locked_count}/{expected_locked}) (0.30 pts)")
                total_score += 0.30
            else:
                print(f"FAIL: Component 2 — Only {locked_count}/{expected_locked} salary cells are locked")
        else:
            print(f"FAIL: Component 2 — Sheet protection not enabled, salary cell lock status irrelevant")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Non-salary cells are unlocked (0.35 points)
    # In initial env, all cells are locked=True (default) — but protection is off.
    # In golden env, non-C-data cells should be explicitly unlocked, and protection is on.
    # Again, gate on sheet protection being on, so initial env scores 0.
    try:
        if ws.protection.sheet:
            unlocked_count = 0
            total_non_salary = 0
            # Check columns A, B, D for rows 1-15 and C1 (header) and C16
            cells_to_check = []
            for row in range(1, 16):
                for col in [1, 2, 4]:  # A, B, D
                    cells_to_check.append((row, col))
            cells_to_check.append((1, 3))  # C1 header should be unlocked

            total_non_salary = len(cells_to_check)
            for row, col in cells_to_check:
                cell = ws.cell(row=row, column=col)
                if not cell.protection.locked:
                    unlocked_count += 1

            # Allow partial credit: if >= 90% are unlocked, full points
            ratio = unlocked_count / total_non_salary if total_non_salary > 0 else 0
            if ratio >= 0.9:
                print(f"PASS: Component 3 — {unlocked_count}/{total_non_salary} non-salary cells are unlocked (ratio={ratio:.2f}) (0.35 pts)")
                total_score += 0.35
            elif ratio >= 0.5:
                partial = round(0.35 * ratio, 2)
                print(f"PARTIAL: Component 3 — {unlocked_count}/{total_non_salary} non-salary cells unlocked (ratio={ratio:.2f}) ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 3 — Only {unlocked_count}/{total_non_salary} non-salary cells are unlocked (ratio={ratio:.2f})")
        else:
            print(f"FAIL: Component 3 — Sheet protection not enabled, unlock status irrelevant")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
