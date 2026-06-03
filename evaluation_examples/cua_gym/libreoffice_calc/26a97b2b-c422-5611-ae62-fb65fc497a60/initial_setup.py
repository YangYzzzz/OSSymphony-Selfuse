"""
Initial Setup: Protect the salary column so only HR managers with the password can edit it.
Task ID: calc_hr_041
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_041'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Confidential ---
    ws = wb.active
    ws.title = 'Confidential'

    # Headers
    headers = ['Employee', 'Department', 'Salary', 'Performance Rating']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic HR data (rows 2-15)
    data = [
        ['Sarah Chen', 'Engineering', 92500, 4.5],
        ['Marcus Johnson', 'Marketing', 68000, 3.8],
        ['Priya Patel', 'Engineering', 105000, 4.9],
        ['David Kim', 'Finance', 78500, 4.1],
        ['Elena Rodriguez', 'Human Resources', 71000, 3.6],
        ['James O\'Brien', 'Sales', 82000, 4.3],
        ['Aisha Mohammed', 'Engineering', 97000, 4.7],
        ['Robert Taylor', 'Operations', 65000, 3.2],
        ['Lisa Wang', 'Marketing', 73500, 4.0],
        ['Carlos Mendez', 'Finance', 88000, 4.4],
        ['Jennifer Park', 'Sales', 76000, 3.9],
        ['Michael Brown', 'Engineering', 110000, 4.8],
        ['Amanda Foster', 'Human Resources', 69500, 3.5],
        ['Thomas Wilson', 'Operations', 72000, 3.7],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Format salary column as currency
    for r in range(2, 16):
        ws.cell(row=r, column=3).number_format = '$#,##0.00'

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20

    # NO sheet protection, NO cell protection overrides
    # This is the pre-task state: everything is default

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
