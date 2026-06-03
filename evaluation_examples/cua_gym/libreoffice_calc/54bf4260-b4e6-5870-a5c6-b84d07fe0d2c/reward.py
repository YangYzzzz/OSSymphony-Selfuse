"""
Reward Script: Apply conditional number format [GREEN][>=1000]#,##0;[RED]#,##0 to B2:B5
Task ID: calc_lf_061
Domain: libreoffice_calc
Scoring:
  Component 1 (0.6): Per-cell number format check on B2:B5 (0.15 each)
  Component 2 (0.4): All 4 cells have correct format AND original values preserved
"""

import os

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_061'
EXPECTED_FORMAT = '[GREEN][>=1000]#,##0;[RED]#,##0'
EXPECTED_VALUES = {
    'B2': 1500,
    'B3': 450,
    'B4': 2300,
    'B5': 870,
}


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet exists
    if 'Targets' not in wb.sheetnames:
        print("FAIL: Sheet 'Targets' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Targets']

    # Component 1: Per-cell number format check on B2:B5 (0.15 points each, 0.6 total)
    # This checks that the conditional number format was applied to each cell.
    # On initial_env, format is 'General' so all fail. On golden_env, format matches.
    format_match_count = 0
    for cell_ref in ['B2', 'B3', 'B4', 'B5']:
        try:
            cell = ws[cell_ref]
            actual_format = cell.number_format
            if actual_format == EXPECTED_FORMAT:
                print(f"PASS: Component 1 — {cell_ref} has correct format '{actual_format}' (0.15 pts)")
                total_score += 0.15
                format_match_count += 1
            else:
                print(f"FAIL: Component 1 — {cell_ref} format is '{actual_format}', expected '{EXPECTED_FORMAT}'")
        except Exception as e:
            print(f"ERROR: Component 1 — {cell_ref}: {e}")

    # Component 2: All 4 cells formatted AND values preserved (0.4 points)
    # This is a compound check: format must be applied AND data must not be corrupted.
    # On initial_env, format_match_count == 0 so this fails.
    # On golden_env, all 4 match and values are intact.
    try:
        if format_match_count == 4:
            values_ok = True
            for cell_ref, expected_val in EXPECTED_VALUES.items():
                actual_val = ws[cell_ref].value
                if actual_val is None or abs(float(actual_val) - expected_val) > 0.01:
                    print(f"FAIL: Component 2 — {cell_ref} value is {actual_val}, expected {expected_val}")
                    values_ok = False
                    break
            if values_ok:
                print(f"PASS: Component 2 — All 4 cells have correct format AND values preserved (0.4 pts)")
                total_score += 0.4
            else:
                print(f"FAIL: Component 2 — Values corrupted despite correct format")
        else:
            print(f"FAIL: Component 2 — Only {format_match_count}/4 cells have correct format (need all 4)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
