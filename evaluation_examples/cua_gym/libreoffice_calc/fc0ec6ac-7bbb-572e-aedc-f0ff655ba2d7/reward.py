"""
Reward Script: Stock Portfolio Tracker
Task ID: calc_wf_071
Domain: libreoffice_calc
Scoring:
  Component 1: Calculated columns (Market Value, Cost Basis, Gain/Loss, Return %, Allocation %) — 0.30
  Component 2: Summary sheet totals — 0.20
  Component 3: Conditional formatting (green gains, red losses) — 0.20
  Component 4: Pie chart for allocation — 0.15
  Component 5: Bar chart for return % by stock — 0.15
"""

import os

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_071'


def verify_task(file_path):
    """
    Verify stock portfolio tracker task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Holdings sheet must exist with base data
    if 'Holdings' not in wb.sheetnames:
        print("FAIL: 'Holdings' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Holdings']

    # ---------------------------------------------------------------------------
    # Component 1: Calculated columns G-K with correct values (0.30 points)
    # These columns do NOT exist in initial_env — they are task-introduced.
    # G=Market Value, H=Cost Basis, I=Gain/Loss, J=Return %, K=Allocation %
    # ---------------------------------------------------------------------------
    try:
        # Check that column headers exist (G1-K1)
        expected_headers = {
            7: 'Market Value', 8: 'Cost Basis', 9: 'Gain/Loss',
            10: 'Return %', 11: 'Allocation %'
        }
        headers_ok = True
        for col, expected in expected_headers.items():
            val = ws.cell(row=1, column=col).value
            if val is None:
                headers_ok = False
                break

        if not headers_ok:
            print("FAIL: Component 1 — calculated column headers (G-K) not found")
        else:
            # Verify computed values for a sample of rows
            # Known ground truth from task context:
            # Market Value = Shares * Current Price
            # Cost Basis = Shares * Purchase Price
            # Gain/Loss = Market Value - Cost Basis
            # Return % = (Gain/Loss / Cost Basis) * 100
            correct_count = 0
            total_checks = 0
            tolerance = 0.5  # allow small rounding differences

            for row in range(2, 12):  # rows 2-11 (10 stocks)
                shares = ws.cell(row=row, column=3).value  # C
                purchase_price = ws.cell(row=row, column=4).value  # D
                current_price = ws.cell(row=row, column=6).value  # F

                if shares is None or purchase_price is None or current_price is None:
                    continue

                expected_mv = shares * current_price
                expected_cb = shares * purchase_price
                expected_gl = expected_mv - expected_cb
                expected_ret = (expected_gl / expected_cb) * 100 if expected_cb != 0 else 0

                # Check Market Value (G)
                mv = ws.cell(row=row, column=7).value
                total_checks += 1
                if mv is not None and abs(float(mv) - expected_mv) <= tolerance:
                    correct_count += 1

                # Check Cost Basis (H)
                cb = ws.cell(row=row, column=8).value
                total_checks += 1
                if cb is not None and abs(float(cb) - expected_cb) <= tolerance:
                    correct_count += 1

                # Check Gain/Loss (I)
                gl = ws.cell(row=row, column=9).value
                total_checks += 1
                if gl is not None and abs(float(gl) - expected_gl) <= tolerance:
                    correct_count += 1

                # Check Return % (J)
                ret = ws.cell(row=row, column=10).value
                total_checks += 1
                if ret is not None and abs(float(ret) - expected_ret) <= tolerance:
                    correct_count += 1

            # Also check Allocation % (K) — requires knowing total market value
            # Sum up all market values from column G
            total_mv = 0
            for row in range(2, 12):
                mv = ws.cell(row=row, column=7).value
                if mv is not None:
                    total_mv += float(mv)

            if total_mv > 0:
                for row in range(2, 12):
                    mv = ws.cell(row=row, column=7).value
                    alloc = ws.cell(row=row, column=11).value
                    total_checks += 1
                    if mv is not None and alloc is not None:
                        expected_alloc = (float(mv) / total_mv) * 100
                        if abs(float(alloc) - expected_alloc) <= tolerance:
                            correct_count += 1

            if total_checks > 0:
                ratio = correct_count / total_checks
                component_score = round(0.30 * ratio, 4)
                print(f"PASS: Component 1 — calculated columns: {correct_count}/{total_checks} values correct ({component_score} pts)")
                total_score += component_score
            else:
                print("FAIL: Component 1 — no calculated values found in columns G-K")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ---------------------------------------------------------------------------
    # Component 2: Summary sheet totals (0.20 points)
    # Initial Summary sheet has only "Portfolio Summary" in A1, no totals.
    # Golden has Total Market Value, Total Cost Basis, Overall Gain/Loss, Return %
    # ---------------------------------------------------------------------------
    try:
        if 'Summary' not in wb.sheetnames:
            print("FAIL: Component 2 — 'Summary' sheet not found")
        else:
            ws_sum = wb['Summary']
            # Check for presence of summary values (not just the header)
            # We look for at least 3 of 4 key summary metrics as numeric values
            summary_values_found = 0

            # Scan summary sheet for numeric values in B column (or wherever they are)
            for row in ws_sum.iter_rows(min_row=2, max_row=20, min_col=1, max_col=5, values_only=False):
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        summary_values_found += 1

            if summary_values_found >= 3:
                # Additionally verify approximate correctness of totals
                # Expected: Total MV ~255876.75, Total CB ~143824.75, GL ~112052, Ret ~77.91
                expected_total_mv = 255876.75
                expected_total_cb = 143824.75
                expected_gl = 112052.0
                expected_ret = 77.91

                # Find values in the sheet
                found_mv = False
                found_cb = False
                found_gl = False
                found_ret = False

                for row in ws_sum.iter_rows(min_row=2, max_row=20, min_col=1, max_col=5, values_only=False):
                    for cell in row:
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            v = float(cell.value)
                            if abs(v - expected_total_mv) < 5:
                                found_mv = True
                            elif abs(v - expected_total_cb) < 5:
                                found_cb = True
                            elif abs(v - expected_gl) < 5:
                                found_gl = True
                            elif abs(v - expected_ret) < 2:
                                found_ret = True

                matches = sum([found_mv, found_cb, found_gl, found_ret])
                if matches >= 3:
                    print(f"PASS: Component 2 — Summary totals correct ({matches}/4 metrics found) (0.20 pts)")
                    total_score += 0.20
                elif matches >= 2:
                    print(f"PARTIAL: Component 2 — Summary totals partially correct ({matches}/4 metrics) (0.10 pts)")
                    total_score += 0.10
                else:
                    print(f"FAIL: Component 2 — Summary totals mostly wrong ({matches}/4 metrics)")
            else:
                print(f"FAIL: Component 2 — insufficient summary values found ({summary_values_found})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---------------------------------------------------------------------------
    # Component 3: Conditional formatting on Return % column (0.20 points)
    # Initial has NO conditional formatting. Golden has green for >0, red for <0.
    # ---------------------------------------------------------------------------
    try:
        cf_rules = list(ws.conditional_formatting)
        if len(cf_rules) == 0:
            print("FAIL: Component 3 — no conditional formatting rules found")
        else:
            # Look for rules that apply to the J column (Return %) or I column (Gain/Loss)
            has_green_rule = False
            has_red_rule = False

            for cf in cf_rules:
                cf_range = str(cf)
                for rule in cf.rules:
                    if rule.type == 'cellIs':
                        dxf = rule.dxf
                        fill_color = None
                        if dxf and dxf.fill and dxf.fill.fgColor:
                            fill_color = dxf.fill.fgColor.rgb

                        # Check for green (positive) rule
                        if rule.operator in ('greaterThan', 'greaterThanOrEqual') and fill_color:
                            green_colors = ['FF00B050', 'FF00FF00', 'FF92D050', 'FF00B04F']
                            if fill_color in green_colors or (fill_color and '00' in fill_color[2:4] and fill_color[4:6] > 'A0'):
                                has_green_rule = True

                        # Check for red (negative) rule
                        if rule.operator in ('lessThan', 'lessThanOrEqual') and fill_color:
                            red_colors = ['FFFF0000', 'FFFF4444', 'FFFF6666']
                            if fill_color in red_colors or (fill_color and fill_color[2:4] > 'C0' and fill_color[4:6] < '40'):
                                has_red_rule = True

            if has_green_rule and has_red_rule:
                print("PASS: Component 3 — conditional formatting: green for gains, red for losses (0.20 pts)")
                total_score += 0.20
            elif has_green_rule or has_red_rule:
                partial = "green" if has_green_rule else "red"
                print(f"PARTIAL: Component 3 — only {partial} conditional formatting found (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 3 — conditional formatting rules exist but no green/red pattern found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ---------------------------------------------------------------------------
    # Component 4: Pie chart for allocation (0.15 points)
    # Initial has 0 charts. Golden has a PieChart titled "Portfolio Allocation".
    # ---------------------------------------------------------------------------
    try:
        from openpyxl.chart import PieChart as PieChartType

        # Search all sheets for pie charts
        pie_found = False
        for sn in wb.sheetnames:
            sheet = wb[sn]
            for chart in sheet._charts:
                if isinstance(chart, PieChartType):
                    pie_found = True
                    print(f"PASS: Component 4 — Pie chart found in sheet '{sn}' (0.15 pts)")
                    break
            if pie_found:
                break

        if not pie_found:
            print("FAIL: Component 4 — no pie chart found in any sheet")
        else:
            total_score += 0.15
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ---------------------------------------------------------------------------
    # Component 5: Bar chart for return % by stock (0.15 points)
    # Initial has 0 charts. Golden has a BarChart titled "Return % by Stock".
    # ---------------------------------------------------------------------------
    try:
        from openpyxl.chart import BarChart as BarChartType

        bar_found = False
        for sn in wb.sheetnames:
            sheet = wb[sn]
            for chart in sheet._charts:
                if isinstance(chart, BarChartType):
                    bar_found = True
                    print(f"PASS: Component 5 — Bar chart found in sheet '{sn}' (0.15 pts)")
                    break
            if bar_found:
                break

        if not bar_found:
            print("FAIL: Component 5 — no bar chart found in any sheet")
        else:
            total_score += 0.15
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved GUI state before verifying
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
