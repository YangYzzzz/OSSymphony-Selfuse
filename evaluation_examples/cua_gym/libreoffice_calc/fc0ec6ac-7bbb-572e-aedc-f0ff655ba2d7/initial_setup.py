"""
Initial Setup: Stock Portfolio Tracker
Task ID: calc_wf_071
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_071'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Holdings ---
    ws = wb.active
    ws.title = 'Holdings'

    headers = ['Symbol', 'Name', 'Shares', 'Purchase Price', 'Purchase Date', 'Current Price']
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # 10 realistic stock holdings
    data = [
        ['AAPL',  'Apple Inc.',              150, 142.50, '2023-03-15', 178.25],
        ['MSFT',  'Microsoft Corp.',          80, 285.00, '2023-06-20', 415.60],
        ['GOOGL', 'Alphabet Inc.',            45, 108.75, '2023-01-10', 155.30],
        ['AMZN',  'Amazon.com Inc.',          60, 98.20,  '2023-04-05', 185.40],
        ['NVDA',  'NVIDIA Corp.',            100, 220.00, '2023-08-12', 875.50],
        ['TSLA',  'Tesla Inc.',               35, 245.00, '2023-02-28', 172.80],
        ['META',  'Meta Platforms Inc.',       55, 180.50, '2023-05-18', 505.75],
        ['JPM',   'JPMorgan Chase & Co.',     90, 138.25, '2023-07-01', 198.40],
        ['JNJ',   'Johnson & Johnson',       120, 165.80, '2022-11-22', 156.20],
        ['V',     'Visa Inc.',                70, 228.90, '2023-09-14', 282.35],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = header_border
            if c in (4, 6):  # Purchase Price, Current Price
                cell.number_format = '$#,##0.00'
            elif c == 5:  # Date
                cell.alignment = Alignment(horizontal="center")

    # Set column widths for readability
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    # --- Sheet 2: Summary (empty placeholder - agent needs to populate) ---
    ws2 = wb.create_sheet('Summary')
    ws2['A1'] = 'Portfolio Summary'
    ws2['A1'].font = Font(bold=True, size=14)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
