"""
Initial Setup: Calculate warehouse utilization percentage for each facility
Task ID: calc_ops_010
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time

subprocess.run(['pip3', 'install', 'openpyxl'], capture_output=True)
import openpyxl
from openpyxl.styles import Font, Alignment, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_010'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Capacity'

    # Headers
    ws['A1'] = 'Warehouse'
    ws['B1'] = 'Total Capacity (pallets)'
    ws['C1'] = 'Used (pallets)'
    ws['D1'] = 'Utilization %'

    # Bold headers
    header_font = Font(bold=True)
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}1'].font = header_font

    # Data rows
    data = [
        ['WH-Alpha', 5000, 4250],
        ['WH-Beta',  3000, 2100],
        ['WH-Gamma', 8000, 7600],
        ['WH-Delta', 4500, 2700],
    ]
    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        ws.cell(row=r, column=2, value=row_data[1])
        ws.cell(row=r, column=3, value=row_data[2])

    # D2:D5 are intentionally left EMPTY - that's the task for the agent

    # Column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
