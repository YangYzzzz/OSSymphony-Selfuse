"""
Reward Script: Warehouse utilization percentage formula verification
Task ID: calc_ops_010
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): D2:D5 all contain formulas (not empty/static)
  Component 2 (0.35): Formulas use division of C/B (=C{n}/B{n})
  Component 3 (0.25): Computed values match expected percentages (85%, 70%, 95%, 60%)
  Component 4 (0.15): D2:D5 formatted as percentage
"""

import os
import re
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_010'

EXPECTED_VALUES = {
    2: 0.85,  # WH-Alpha: 4250/5000 = 85%
    3: 0.70,  # WH-Beta:  2100/3000 = 70%
    4: 0.95,  # WH-Gamma: 7600/8000 = 95%
    5: 0.60,  # WH-Delta: 2700/4500 = 60%
}


def persist_app_state(domain: str):
    """Save any unsaved LibreOffice edits via Ctrl+S."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook with formulas
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb["Capacity"]

    # Component 1: D2:D5 all contain formulas — 0.25 points
    try:
        formula_count = 0
        for row in range(2, 6):
            cell_val = ws.cell(row=row, column=4).value
            if cell_val is not None and isinstance(cell_val, str) and cell_val.startswith("="):
                formula_count += 1
            else:
                print(f"FAIL: Component 1 -- D{row} does not contain a formula, found: {repr(cell_val)}")

        if formula_count == 4:
            print(f"PASS: Component 1 -- All 4 cells D2:D5 contain formulas (0.25 pts)")
            total_score += 0.25
        elif formula_count > 0:
            partial = round(0.25 * formula_count / 4, 2)
            print(f"PARTIAL: Component 1 -- {formula_count}/4 cells contain formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 -- No formulas found in D2:D5")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Formulas use C/B division pattern — 0.35 points
    try:
        correct_formulas = 0
        for row in range(2, 6):
            cell_val = ws.cell(row=row, column=4).value
            if cell_val is not None and isinstance(cell_val, str):
                normalized = cell_val.upper().replace(" ", "")
                # Accept =C{n}/B{n} pattern
                pattern = f"=C{row}/B{row}"
                if normalized == pattern:
                    correct_formulas += 1
                else:
                    # Also accept equivalent forms like =(C2/B2)
                    alt_pattern = re.compile(
                        rf'^=\(?C{row}\s*/\s*B{row}\)?$', re.IGNORECASE
                    )
                    if alt_pattern.match(cell_val.strip()):
                        correct_formulas += 1
                    else:
                        print(f"FAIL: Component 2 -- D{row} formula {repr(cell_val)} does not match =C{row}/B{row}")
            else:
                print(f"FAIL: Component 2 -- D{row} is not a formula: {repr(cell_val)}")

        if correct_formulas == 4:
            print(f"PASS: Component 2 -- All 4 formulas correctly divide Used/Capacity (0.35 pts)")
            total_score += 0.35
        elif correct_formulas > 0:
            partial = round(0.35 * correct_formulas / 4, 2)
            print(f"PARTIAL: Component 2 -- {correct_formulas}/4 formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- No correct division formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Computed values match expected percentages — 0.25 points
    try:
        wb_data = openpyxl.load_workbook(file_path, data_only=True)
        ws_data = wb_data["Capacity"]

        correct_values = 0
        for row, expected in EXPECTED_VALUES.items():
            cached_val = ws_data.cell(row=row, column=4).value
            if cached_val is not None:
                try:
                    numeric_val = float(cached_val)
                    if abs(numeric_val - expected) < 0.01:
                        correct_values += 1
                    else:
                        print(f"FAIL: Component 3 -- D{row} computed value {numeric_val}, expected {expected}")
                except (ValueError, TypeError):
                    print(f"FAIL: Component 3 -- D{row} value {repr(cached_val)} is not numeric")
            else:
                # Cached value may be None; try manual evaluation
                formula = ws.cell(row=row, column=4).value
                if formula and isinstance(formula, str) and "/" in formula:
                    c_val = ws.cell(row=row, column=3).value
                    b_val = ws.cell(row=row, column=2).value
                    if c_val is not None and b_val is not None and b_val != 0:
                        manual_result = float(c_val) / float(b_val)
                        if abs(manual_result - expected) < 0.01:
                            correct_values += 1
                            print(f"INFO: D{row} manual evaluation = {manual_result} (matches expected {expected})")
                        else:
                            print(f"FAIL: Component 3 -- D{row} manual evaluation {manual_result}, expected {expected}")
                    else:
                        print(f"FAIL: Component 3 -- D{row} cannot evaluate: B={b_val}, C={c_val}")
                else:
                    print(f"FAIL: Component 3 -- D{row} cached is None, no formula to evaluate")

        if correct_values == 4:
            print(f"PASS: Component 3 -- All 4 utilization values correct: 85%, 70%, 95%, 60% (0.25 pts)")
            total_score += 0.25
        elif correct_values > 0:
            partial = round(0.25 * correct_values / 4, 2)
            print(f"PARTIAL: Component 3 -- {correct_values}/4 values correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 -- No correct utilization values")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: D2:D5 formatted as percentage — 0.15 points
    try:
        pct_count = 0
        for row in range(2, 6):
            fmt = ws.cell(row=row, column=4).number_format
            if fmt and ('%' in fmt or 'percent' in fmt.lower()):
                pct_count += 1
            else:
                print(f"FAIL: Component 4 -- D{row} format is {repr(fmt)}, not percentage")

        if pct_count == 4:
            print(f"PASS: Component 4 -- All 4 cells formatted as percentage (0.15 pts)")
            total_score += 0.15
        elif pct_count > 0:
            partial = round(0.15 * pct_count / 4, 2)
            print(f"PARTIAL: Component 4 -- {pct_count}/4 cells formatted as percentage ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 -- No percentage formatting found in D2:D5")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    final_score = round(min(total_score, 1.0), 1)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist any unsaved edits before verification
persist_app_state("libreoffice_calc")

# Run verification
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
