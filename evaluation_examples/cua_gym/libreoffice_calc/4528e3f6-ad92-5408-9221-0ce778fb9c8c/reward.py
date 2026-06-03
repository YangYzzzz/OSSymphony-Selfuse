"""
Reward Script: Accounts Receivable Aging Report
Task ID: calc_wf_062
Domain: libreoffice_calc
Scoring:
  Component 1: Days Outstanding column G formulas (0.20)
  Component 2: Aging Bucket column H formulas (0.20)
  Component 3: Summary table with SUMIFS (0.20)
  Component 4: Stacked bar chart with 5 series (0.20)
  Component 5: Conditional formatting rules (0.20)
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_062'


def persist_app_state(domain):
    """Best-effort save of any open LibreOffice document."""
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print(f"PERSIST: ctrl+s sent for {domain}")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Ensure 'Invoices' sheet exists
    if 'Invoices' not in wb.sheetnames:
        print("CRITICAL: 'Invoices' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Invoices']

    # ---------------------------------------------------------------
    # Component 1: Days Outstanding column G with TODAY()-D formulas (0.20 pts)
    # Initial file has no column G formulas; golden has =TODAY()-D# in G2:G41
    # ---------------------------------------------------------------
    try:
        g_formula_count = 0
        for row in range(2, 42):  # rows 2-41 (40 invoices)
            val = ws.cell(row=row, column=7).value  # column G
            if val is not None and isinstance(val, str):
                # Check for TODAY()-D pattern
                normalized = val.upper().replace(" ", "")
                if "TODAY()" in normalized and f"-D{row}" in normalized:
                    g_formula_count += 1
        if g_formula_count >= 35:
            # At least 35 out of 40 have proper formulas
            print(f"PASS: Component 1 - Days Outstanding formulas found: {g_formula_count}/40 (0.20 pts)")
            total_score += 0.20
        elif g_formula_count >= 20:
            partial = 0.10
            print(f"PARTIAL: Component 1 - Days Outstanding formulas found: {g_formula_count}/40 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - Days Outstanding formulas found: {g_formula_count}/40")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # ---------------------------------------------------------------
    # Component 2: Aging Bucket column H with nested IF formulas (0.20 pts)
    # Initial file has no column H; golden has nested IF in H2:H41
    # ---------------------------------------------------------------
    try:
        h_formula_count = 0
        for row in range(2, 42):
            val = ws.cell(row=row, column=8).value  # column H
            if val is not None and isinstance(val, str):
                normalized = val.upper().replace(" ", "")
                # Check for nested IF with aging bucket keywords
                if "IF(" in normalized and ("CURRENT" in normalized or "1-30" in normalized or "90+" in normalized):
                    h_formula_count += 1
        if h_formula_count >= 35:
            print(f"PASS: Component 2 - Aging Bucket formulas found: {h_formula_count}/40 (0.20 pts)")
            total_score += 0.20
        elif h_formula_count >= 20:
            partial = 0.10
            print(f"PARTIAL: Component 2 - Aging Bucket formulas found: {h_formula_count}/40 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 - Aging Bucket formulas found: {h_formula_count}/40")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # ---------------------------------------------------------------
    # Component 3: Summary table with SUMIFS by customer and Grand Total (0.20 pts)
    # Golden has a summary section starting around row 44 with headers and SUMIFS
    # Initial file has nothing below row 41
    # ---------------------------------------------------------------
    try:
        summary_score = 0.0

        # Check for summary header row containing bucket labels
        found_summary_header = False
        summary_header_row = None
        for row in range(42, 60):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val and isinstance(cell_val, str) and "customer" in cell_val.lower():
                # Check if adjacent columns have bucket labels
                b_val = ws.cell(row=row, column=2).value
                if b_val and isinstance(b_val, str) and "current" in b_val.lower():
                    found_summary_header = True
                    summary_header_row = row
                    break

        if found_summary_header:
            summary_score += 0.05
            print(f"  Sub-check: Summary header row found at row {summary_header_row}")

            # Check for SUMIFS formulas in the customer data rows
            sumifs_count = 0
            customer_count = 0
            for row in range(summary_header_row + 1, summary_header_row + 15):
                a_val = ws.cell(row=row, column=1).value
                if a_val and isinstance(a_val, str) and a_val.strip():
                    if a_val.strip().lower() == "grand total":
                        continue
                    customer_count += 1
                    # Check columns B-F for SUMIFS
                    for col in range(2, 7):
                        cval = ws.cell(row=row, column=col).value
                        if cval and isinstance(cval, str) and "SUMIFS" in cval.upper():
                            sumifs_count += 1
                            break  # at least one SUMIFS per customer is enough

            if customer_count >= 8:
                summary_score += 0.05
                print(f"  Sub-check: {customer_count} customers found in summary")

            if sumifs_count >= 8:
                summary_score += 0.05
                print(f"  Sub-check: {sumifs_count} customers have SUMIFS formulas")

            # Check for Grand Total row with SUM formulas
            grand_total_found = False
            for row in range(summary_header_row + 1, summary_header_row + 15):
                a_val = ws.cell(row=row, column=1).value
                if a_val and isinstance(a_val, str) and "total" in a_val.lower():
                    b_val = ws.cell(row=row, column=2).value
                    if b_val and isinstance(b_val, str) and "SUM(" in b_val.upper():
                        grand_total_found = True
                        break
            if grand_total_found:
                summary_score += 0.05
                print(f"  Sub-check: Grand Total row with SUM formulas found")
            else:
                print(f"  Sub-check FAIL: Grand Total row not found or missing SUM formulas")
        else:
            print(f"  FAIL: Summary header row not found")

        if summary_score >= 0.19:
            print(f"PASS: Component 3 - Summary table ({summary_score:.2f} pts)")
        elif summary_score > 0:
            print(f"PARTIAL: Component 3 - Summary table ({summary_score:.2f} pts)")
        else:
            print(f"FAIL: Component 3 - Summary table not found")
        total_score += summary_score

    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # ---------------------------------------------------------------
    # Component 4: Stacked bar chart with 5 series (0.20 pts)
    # Initial file has 0 charts; golden has 1 stacked bar chart with 5 series
    # ---------------------------------------------------------------
    try:
        charts = ws._charts
        if len(charts) == 0:
            print(f"FAIL: Component 4 - No charts found")
        else:
            chart = charts[0]
            chart_score = 0.0

            # Check chart exists and is a bar/column type
            from openpyxl.chart import BarChart
            if isinstance(chart, BarChart):
                chart_score += 0.05
                print(f"  Sub-check: Chart is BarChart type")
            else:
                print(f"  Sub-check FAIL: Chart is {type(chart).__name__}, expected BarChart")

            # Check it's stacked
            grouping = getattr(chart, 'grouping', None)
            if grouping and 'stacked' in str(grouping).lower():
                chart_score += 0.05
                print(f"  Sub-check: Chart grouping is {grouping}")
            else:
                print(f"  Sub-check FAIL: Chart grouping is {grouping}, expected stacked")

            # Check 5 series (one per aging bucket)
            series_count = len(chart.series)
            if series_count == 5:
                chart_score += 0.05
                print(f"  Sub-check: Chart has {series_count} series (5 aging buckets)")
            elif series_count >= 3:
                chart_score += 0.03
                print(f"  Sub-check PARTIAL: Chart has {series_count} series, expected 5")
            else:
                print(f"  Sub-check FAIL: Chart has {series_count} series, expected 5")

            # Check chart has a title
            if chart.title is not None:
                chart_score += 0.05
                print(f"  Sub-check: Chart has a title")
            else:
                print(f"  Sub-check FAIL: Chart has no title")

            if chart_score >= 0.19:
                print(f"PASS: Component 4 - Stacked bar chart ({chart_score:.2f} pts)")
            elif chart_score > 0:
                print(f"PARTIAL: Component 4 - Chart ({chart_score:.2f} pts)")
            else:
                print(f"FAIL: Component 4 - Chart issues")
            total_score += chart_score
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # ---------------------------------------------------------------
    # Component 5: Conditional formatting for overdue amounts (0.20 pts)
    # Initial file has 0 CF rules; golden has CF on G2:G41 and C2:C41
    # Orange (61-90 days) and Red (90+ days)
    # ---------------------------------------------------------------
    try:
        cf_rules = list(ws.conditional_formatting)
        if len(cf_rules) == 0:
            print(f"FAIL: Component 5 - No conditional formatting rules found")
        else:
            cf_score = 0.0

            # Check for any CF rule related to 61-90 (orange) and 90+ (red)
            found_orange = False
            found_red = False

            for cf in cf_rules:
                for rule in cf.rules:
                    formulas = getattr(rule, 'formula', []) or []
                    op = getattr(rule, 'operator', None)

                    # Check fill color
                    fill_color = None
                    if rule.dxf and rule.dxf.fill:
                        try:
                            fill_color = rule.dxf.fill.fgColor.rgb
                        except:
                            pass

                    # Orange rule: 61-90 range
                    if fill_color and 'FF99' in fill_color.upper() or fill_color == 'FFFF9900':
                        found_orange = True

                    # Red rule: >90
                    if fill_color and fill_color in ('FFFF0000', 'FF0000'):
                        found_red = True

                    # Also check by formula content
                    for f in formulas:
                        f_str = str(f)
                        if '61' in f_str and '90' in f_str:
                            found_orange = True
                        if ('90' in f_str or '>90' in f_str) and '61' not in f_str:
                            found_red = True

                    # Check operator-based rules
                    if op == 'between' and formulas:
                        f_strs = [str(x) for x in formulas]
                        if '61' in f_strs and '90' in f_strs:
                            found_orange = True
                    if op == 'greaterThan' and formulas:
                        f_strs = [str(x) for x in formulas]
                        if '90' in f_strs:
                            found_red = True

            if found_orange:
                cf_score += 0.10
                print(f"  Sub-check: Orange conditional formatting (61-90 days) found")
            else:
                print(f"  Sub-check FAIL: Orange conditional formatting (61-90 days) not found")

            if found_red:
                cf_score += 0.10
                print(f"  Sub-check: Red conditional formatting (90+ days) found")
            else:
                print(f"  Sub-check FAIL: Red conditional formatting (90+ days) not found")

            total_cf_rules = sum(len(cf.rules) for cf in cf_rules)
            print(f"  Total CF rules: {total_cf_rules} across {len(cf_rules)} ranges")

            if cf_score >= 0.19:
                print(f"PASS: Component 5 - Conditional formatting ({cf_score:.2f} pts)")
            elif cf_score > 0:
                print(f"PARTIAL: Component 5 - Conditional formatting ({cf_score:.2f} pts)")
            else:
                print(f"FAIL: Component 5 - Conditional formatting issues")
            total_score += cf_score
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
