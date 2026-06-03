"""
Initial Setup: Accounts Receivable Aging Report
Task ID: calc_wf_062
Domain: libreoffice_calc

Creates an Invoices sheet with 40 invoices across 10 customers.
Raw data only — no formulas, no aging buckets, no summary, no charts,
no conditional formatting.
"""

import os
import shlex
import subprocess
import time
from datetime import date, timedelta
import random

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_062'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)  # reproducible data

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoices"

    # --- Header row ---
    headers = [
        "Invoice #", "Customer", "Amount", "Invoice Date",
        "Payment Terms", "Due Date"
    ]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="000000")
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # --- Customer names (10 customers) ---
    customers = [
        "Meridian Consulting Group",
        "Apex Technologies Inc.",
        "Brightpath Solutions LLC",
        "CoreVista Partners",
        "Dunmore & Associates",
        "Evergreen Supply Co.",
        "Frontier Digital Services",
        "Greenleaf Manufacturing",
        "Harbor Point Logistics",
        "Ironclad Security Systems",
    ]

    payment_terms_options = ["Net 15", "Net 30", "Net 45", "Net 60"]
    term_days = {"Net 15": 15, "Net 30": 30, "Net 45": 45, "Net 60": 60}

    today = date.today()

    # Generate 40 invoices spread across 10 customers (4 per customer)
    invoices = []
    inv_num = 1001
    for cust in customers:
        for _ in range(4):
            days_ago = random.randint(1, 120)
            inv_date = today - timedelta(days=days_ago)
            amount = round(random.uniform(500, 15000), 2)
            terms = random.choice(payment_terms_options)
            due_date = inv_date + timedelta(days=term_days[terms])
            invoices.append((
                f"INV-{inv_num}", cust, amount, inv_date, terms, due_date
            ))
            inv_num += 1

    # Shuffle to make it realistic (not grouped by customer)
    random.shuffle(invoices)

    # Write invoice data
    for r, row_data in enumerate(invoices, 2):
        inv_id, cust, amount, inv_date, terms, due_date = row_data
        ws.cell(row=r, column=1, value=inv_id)
        ws.cell(row=r, column=2, value=cust)

        amt_cell = ws.cell(row=r, column=3, value=amount)
        amt_cell.number_format = '$#,##0.00'

        date_cell = ws.cell(row=r, column=4, value=inv_date)
        date_cell.number_format = 'yyyy-mm-dd'

        ws.cell(row=r, column=5, value=terms)

        due_cell = ws.cell(row=r, column=6, value=due_date)
        due_cell.number_format = 'yyyy-mm-dd'

    # --- Column widths ---
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f"Initial file created: {OUTPUT}")

    # Open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print("GUI_READY: launched LibreOffice Calc with DISPLAY=:0")


create_initial()
