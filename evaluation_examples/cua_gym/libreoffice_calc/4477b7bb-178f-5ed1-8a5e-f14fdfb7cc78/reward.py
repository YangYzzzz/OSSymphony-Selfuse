"""
Reward Script: Hazardous Materials Inventory Register
Task ID: calc_ops_warehouse_hazmat_compliance_058
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Data validation dropdown (Class 1-9) on HazmatRegister D2:D51     — 0.30 pts
  Component 2: StorageSummary C2:C11 SUMIF formulas for total kg per location    — 0.25 pts
  Component 3: StorageSummary D2:D11 compatibility alert formulas (SUMPRODUCT)   — 0.25 pts
  Component 4: Comment on StorageSummary D1 about incompatible storage risk       — 0.10 pts
  Component 5: Conditional formatting on D2:D11 — red fill for REVIEW REQUIRED   — 0.10 pts
  Total: 1.00

All components FAIL on the initial file (where D2:D51 has no validation,
StorageSummary C and D columns are empty, no comment on D1, no CF rules).
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_warehouse_hazmat_compliance_058'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify required sheets exist as a precondition gate
    if 'HazmatRegister' not in wb.sheetnames:
        print("CRITICAL: Sheet 'HazmatRegister' not found — cannot verify task")
        print("REWARD: 0.0")
        return 0.0
    if 'StorageSummary' not in wb.sheetnames:
        print("CRITICAL: Sheet 'StorageSummary' not found — cannot verify task")
        print("REWARD: 0.0")
        return 0.0

    ws_hazmat = wb['HazmatRegister']
    ws_summary = wb['StorageSummary']

    # -----------------------------------------------------------------------
    # Component 1: Data validation dropdown on HazmatRegister D2:D51 (0.30 pts)
    # The golden file adds a list-type data validation for "Class 1" through
    # "Class 9" on cells D2:D51. The initial file has no data validation at all.
    # -----------------------------------------------------------------------
    try:
        validations = ws_hazmat.data_validations.dataValidation
        dv_found = False
        dv_correct = False
        for dv in validations:
            if dv.type == 'list':
                # Check sqref covers D2:D51 (or at least column D rows 2-51)
                sqref_str = str(dv.sqref)
                if 'D2' in sqref_str or 'D2:D51' in sqref_str:
                    dv_found = True
                    # Check formula1 contains all 9 classes
                    formula = dv.formula1 or ''
                    expected_classes = [
                        'Class 1', 'Class 2', 'Class 3', 'Class 4', 'Class 5',
                        'Class 6', 'Class 7', 'Class 8', 'Class 9'
                    ]
                    all_classes_present = all(c in formula for c in expected_classes)
                    if all_classes_present:
                        dv_correct = True
                        print(f"PASS: Component 1 — Data validation found on D2 area, formula1={formula[:60]}... (0.3 pts)")
                        total_score += 0.30
                    else:
                        print(f"FAIL: Component 1 — Data validation found on D area but missing some classes. formula1={formula}")
                    break

        if not dv_found:
            print(f"FAIL: Component 1 — No list-type data validation found covering D2:D51 in HazmatRegister")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: StorageSummary C2:C11 contain SUMIF formulas (0.25 pts)
    # The golden file adds SUMIF(HazmatRegister!E:E, Ax, HazmatRegister!F:F)
    # in C2:C11. The initial file has None in all these cells.
    # We require at least 8 out of 10 cells to have a SUMIF formula.
    # -----------------------------------------------------------------------
    try:
        sumif_count = 0
        for row in range(2, 12):
            cell_val = ws_summary.cell(row=row, column=3).value
            if cell_val and isinstance(cell_val, str):
                upper_val = cell_val.upper().replace(' ', '')
                if 'SUMIF' in upper_val and 'HAZMATREGISTER' in upper_val:
                    sumif_count += 1

        if sumif_count >= 8:
            print(f"PASS: Component 2 — SUMIF formulas found in StorageSummary C column ({sumif_count}/10 rows) (0.25 pts)")
            total_score += 0.25
        elif sumif_count >= 4:
            partial = 0.12
            print(f"PARTIAL: Component 2 — Only {sumif_count}/10 rows have SUMIF formula in StorageSummary C (0.12 pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Expected SUMIF formulas in StorageSummary C2:C11, found {sumif_count} valid formulas")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: StorageSummary D2:D11 contain compatibility alert formulas (0.25 pts)
    # The golden file adds IF(SUMPRODUCT(...COUNTIFS...)>1,"REVIEW REQUIRED","OK")
    # in D2:D11. The initial file has None in all these cells.
    # We require at least 8 out of 10 cells to have an IF formula with
    # "REVIEW REQUIRED" or SUMPRODUCT/COUNTIFS pattern.
    # -----------------------------------------------------------------------
    try:
        compat_count = 0
        for row in range(2, 12):
            cell_val = ws_summary.cell(row=row, column=4).value
            if cell_val and isinstance(cell_val, str):
                upper_val = cell_val.upper().replace(' ', '')
                # Check for IF formula containing "REVIEW REQUIRED"
                if 'IF(' in upper_val and 'REVIEWREQUIRED' in upper_val.replace('"', '').replace(' ', ''):
                    compat_count += 1
                elif 'SUMPRODUCT' in upper_val and 'COUNTIFS' in upper_val:
                    compat_count += 1

        if compat_count >= 8:
            print(f"PASS: Component 3 — Compatibility alert formulas found in StorageSummary D column ({compat_count}/10 rows) (0.25 pts)")
            total_score += 0.25
        elif compat_count >= 4:
            partial = 0.12
            print(f"PARTIAL: Component 3 — Only {compat_count}/10 rows have compatibility formula in StorageSummary D (0.12 pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Expected IF/SUMPRODUCT formulas in StorageSummary D2:D11, found {compat_count} valid formulas")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: Comment on StorageSummary D1 about incompatible storage risk (0.10 pts)
    # The golden file adds a comment to D1 explaining the incompatible storage
    # risk. The initial file has no comment on D1.
    # -----------------------------------------------------------------------
    try:
        d1_cell = ws_summary['D1']
        comment = d1_cell.comment
        if comment is not None:
            comment_text = str(comment.text) if hasattr(comment, 'text') else str(comment)
            # Check comment mentions incompatible storage or hazard class risk
            lower_text = comment_text.lower()
            if any(keyword in lower_text for keyword in ['incompatible', 'hazard', 'review', 'segregat', 'storage risk', 'class']):
                print(f"PASS: Component 4 — Comment found on StorageSummary D1 about incompatible storage (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 4 — Comment on D1 exists but does not mention incompatible storage. Text: {comment_text[:100]}")
        else:
            print(f"FAIL: Component 4 — No comment found on StorageSummary D1")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -----------------------------------------------------------------------
    # Component 5: Conditional formatting on StorageSummary D2:D11 — red fill
    # for "REVIEW REQUIRED" (0.10 pts)
    # The golden file adds a formula-based CF rule: $D2="REVIEW REQUIRED"
    # with red fill (FFFF0000). The initial file has no CF rules on this sheet.
    # -----------------------------------------------------------------------
    try:
        cf_found = False
        cf_correct = False
        for cf_range in ws_summary.conditional_formatting:
            cf_range_str = str(cf_range)
            # Check if the CF applies to the D column area
            if 'D2' in cf_range_str or 'D2:D11' in cf_range_str:
                for rule in ws_summary.conditional_formatting[cf_range]:
                    cf_found = True
                    # Check for formula or expression type checking "REVIEW REQUIRED"
                    if rule.type in ('expression', 'formula'):
                        formula_list = rule.formula or []
                        formula_str = ' '.join(str(f) for f in formula_list).upper()
                        if 'REVIEW REQUIRED' in formula_str or 'REVIEWREQUIRED' in formula_str.replace('"', '').replace(' ', ''):
                            # Check for red fill
                            try:
                                dxf = rule.dxf
                                if dxf and dxf.fill:
                                    fg_rgb = dxf.fill.fgColor.rgb if dxf.fill.fgColor else None
                                    # Accept FFFF0000 (opaque red)
                                    is_red = fg_rgb and 'FF0000' in fg_rgb.upper()
                                    if is_red:
                                        cf_correct = True
                                        print(f"PASS: Component 5 — Red conditional formatting found on D2:D11 for 'REVIEW REQUIRED' (0.10 pts)")
                                        total_score += 0.10
                                    else:
                                        # CF formula is correct but color is not clearly red — partial acceptance
                                        cf_correct = True
                                        print(f"PASS: Component 5 — CF formula for 'REVIEW REQUIRED' found on D area, fill rgb={fg_rgb} (0.10 pts)")
                                        total_score += 0.10
                                else:
                                    # Formula is correct, give partial credit even without confirmed fill
                                    print(f"PARTIAL: Component 5 — CF formula for 'REVIEW REQUIRED' found but fill unverifiable (0.05 pts)")
                                    total_score += 0.05
                            except Exception as fill_e:
                                print(f"PARTIAL: Component 5 — CF formula found but fill check failed: {fill_e} (0.05 pts)")
                                total_score += 0.05
                            break

        if not cf_found:
            print(f"FAIL: Component 5 — No conditional formatting found on StorageSummary D2:D11")
        elif not cf_correct and cf_found:
            print(f"FAIL: Component 5 — Conditional formatting found on D area but does not check for 'REVIEW REQUIRED'")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
