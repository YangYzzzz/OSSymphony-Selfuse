"""
Initial Setup: Hazardous Materials Inventory Register
Task ID: calc_ops_warehouse_hazmat_compliance_058
Domain: libreoffice_calc

Creates HazmatRegister sheet with 50 hazmat items (Hazard Class column D empty,
Storage Status column H empty) and StorageSummary sheet with 10 storage locations
(Hazard Classes Present, Qty Total kg, Compatibility Alert columns empty).
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_warehouse_hazmat_compliance_058'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    # ---- Sheet 1: HazmatRegister ----
    ws1 = wb.active
    ws1.title = 'HazmatRegister'

    # Headers: A=Item ID, B=Chemical Name, C=UN Number, D=Hazard Class (EMPTY - needs dropdown),
    #           E=Storage Location, F=Qty kg, G=Compatibility Group, H=Storage Status (EMPTY)
    headers = ['Item ID', 'Chemical Name', 'UN Number', 'Hazard Class',
               'Storage Location', 'Qty kg', 'Compatibility Group', 'Storage Status']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFFFF')

    # 50 hazmat items — Hazard Class (D) is EMPTY, Storage Status (H) is EMPTY
    # 10 storage locations: LOC-A through LOC-J
    # Chemicals spread across locations with realistic UN numbers and compatibility groups
    hazmat_data = [
        # (Item ID, Chemical Name, UN Number, Storage Location, Qty kg, Compatibility Group)
        ('HM-001', 'Ammonium Nitrate',          '0222', 'LOC-A', 500.0,  'A'),
        ('HM-002', 'Trinitrotoluene (TNT)',      '0209', 'LOC-A', 150.0,  'D'),
        ('HM-003', 'Propane Gas',                '1978', 'LOC-B', 320.0,  'N/A'),
        ('HM-004', 'Acetylene',                  '1001', 'LOC-B', 180.0,  'N/A'),
        ('HM-005', 'Butane',                     '1011', 'LOC-B', 240.0,  'N/A'),
        ('HM-006', 'Gasoline',                   '1203', 'LOC-C', 850.0,  'N/A'),
        ('HM-007', 'Ethanol (Denatured)',         '1170', 'LOC-C', 430.0,  'N/A'),
        ('HM-008', 'Acetone',                    '1090', 'LOC-C', 275.0,  'N/A'),
        ('HM-009', 'Sodium Metal',               '1428', 'LOC-D', 90.0,   'N/A'),
        ('HM-010', 'Calcium Carbide',            '1402', 'LOC-D', 120.0,  'N/A'),
        ('HM-011', 'Hydrogen Peroxide 60%',      '2014', 'LOC-E', 200.0,  'N/A'),
        ('HM-012', 'Potassium Permanganate',     '1490', 'LOC-E', 165.0,  'N/A'),
        ('HM-013', 'Sodium Hypochlorite',        '1791', 'LOC-E', 310.0,  'N/A'),
        ('HM-014', 'Chlorine Gas',               '1017', 'LOC-F', 75.0,   'N/A'),
        ('HM-015', 'Hydrochloric Acid 37%',      '1789', 'LOC-F', 400.0,  'N/A'),
        ('HM-016', 'Sulfuric Acid 98%',          '1830', 'LOC-F', 560.0,  'N/A'),
        ('HM-017', 'Sodium Hydroxide',           '1823', 'LOC-F', 380.0,  'N/A'),
        ('HM-018', 'Uranium Hexafluoride',       '2978', 'LOC-G', 50.0,   'N/A'),
        ('HM-019', 'Cesium-137 Source',          '2916', 'LOC-G', 2.5,    'N/A'),
        ('HM-020', 'Iridium-192 Industrial',     '2919', 'LOC-G', 1.2,    'N/A'),
        ('HM-021', 'Arsenic Trioxide',           '1561', 'LOC-H', 45.0,   'N/A'),
        ('HM-022', 'Mercury',                    '2809', 'LOC-H', 80.0,   'N/A'),
        ('HM-023', 'Lead(II) Nitrate',           '1469', 'LOC-H', 130.0,  'N/A'),
        ('HM-024', 'Cyanide Solution 15%',       '3414', 'LOC-H', 95.0,   'N/A'),
        ('HM-025', 'Formaldehyde 37%',           '2209', 'LOC-I', 210.0,  'N/A'),
        ('HM-026', 'Methanol',                   '1230', 'LOC-I', 375.0,  'N/A'),
        ('HM-027', 'Toluene Diisocyanate',       '2078', 'LOC-I', 155.0,  'N/A'),
        ('HM-028', 'Acetic Acid Glacial',        '2789', 'LOC-I', 290.0,  'N/A'),
        ('HM-029', 'Lithium Batteries (bulk)',   '3090', 'LOC-J', 320.0,  'N/A'),
        ('HM-030', 'Magnetized Material',        '2807', 'LOC-J', 55.0,   'N/A'),
        ('HM-031', 'Dry Ice (CO2 solid)',        '1845', 'LOC-B', 400.0,  'N/A'),
        ('HM-032', 'Nitrogen Gas Compressed',    '1066', 'LOC-B', 250.0,  'N/A'),
        ('HM-033', 'Diethyl Ether',              '1155', 'LOC-C', 185.0,  'N/A'),
        ('HM-034', 'Isopropyl Alcohol',          '1219', 'LOC-C', 310.0,  'N/A'),
        ('HM-035', 'Toluene',                    '1294', 'LOC-C', 225.0,  'N/A'),
        ('HM-036', 'Phosphorus White',           '1381', 'LOC-D', 35.0,   'N/A'),
        ('HM-037', 'Potassium Metal',            '2257', 'LOC-D', 28.0,   'N/A'),
        ('HM-038', 'Sodium Peroxide',            '1504', 'LOC-E', 140.0,  'N/A'),
        ('HM-039', 'Perchloric Acid 60%',        '1873', 'LOC-E', 85.0,   'N/A'),
        ('HM-040', 'Bromine',                    '1744', 'LOC-F', 120.0,  'N/A'),
        ('HM-041', 'Nitric Acid 65%',            '2031', 'LOC-F', 340.0,  'N/A'),
        ('HM-042', 'Cobalt-60 Source',           '2916', 'LOC-G', 0.8,    'N/A'),
        ('HM-043', 'Selenium Dioxide',           '2811', 'LOC-H', 60.0,   'N/A'),
        ('HM-044', 'Cadmium Chloride',           '2570', 'LOC-H', 75.0,   'N/A'),
        ('HM-045', 'Phenol',                     '1671', 'LOC-I', 195.0,  'N/A'),
        ('HM-046', 'Styrene Monomer',            '2055', 'LOC-I', 270.0,  'N/A'),
        ('HM-047', 'PCB Waste (Transformers)',   '2315', 'LOC-J', 420.0,  'N/A'),
        ('HM-048', 'Asbestos Fibres',            '2212', 'LOC-J', 180.0,  'N/A'),
        ('HM-049', 'RDX Explosive',              '0072', 'LOC-A', 80.0,   'E'),
        ('HM-050', 'Ammonium Perchlorate',       '0402', 'LOC-A', 200.0,  'D'),
    ]

    # Hazard class assignments per item (D column populated later in golden)
    # For initial file D column is LEFT EMPTY
    for r, row in enumerate(hazmat_data, 2):
        item_id, chem_name, un_num, storage_loc, qty_kg, compat_group = row
        ws1.cell(row=r, column=1, value=item_id)
        ws1.cell(row=r, column=2, value=chem_name)
        ws1.cell(row=r, column=3, value=un_num)
        # Column D (Hazard Class) — LEFT EMPTY: task requires adding dropdown
        ws1.cell(row=r, column=4, value=None)
        ws1.cell(row=r, column=5, value=storage_loc)
        ws1.cell(row=r, column=6, value=qty_kg)
        ws1.cell(row=r, column=7, value=compat_group)
        # Column H (Storage Status) — LEFT EMPTY
        ws1.cell(row=r, column=8, value=None)

    # Column widths
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 32
    ws1.column_dimensions['C'].width = 12
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 16
    ws1.column_dimensions['F'].width = 12
    ws1.column_dimensions['G'].width = 20
    ws1.column_dimensions['H'].width = 18

    # ---- Sheet 2: StorageSummary ----
    ws2 = wb.create_sheet('StorageSummary')

    # Headers: A=Location, B=Hazard Classes Present (EMPTY), C=Qty Total kg (EMPTY), D=Compatibility Alert (EMPTY)
    sum_headers = ['Location', 'Hazard Classes Present', 'Qty Total kg', 'Compatibility Alert']
    for col, h in enumerate(sum_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFFFF')

    # 10 storage locations
    locations = ['LOC-A', 'LOC-B', 'LOC-C', 'LOC-D', 'LOC-E',
                 'LOC-F', 'LOC-G', 'LOC-H', 'LOC-I', 'LOC-J']
    for r, loc in enumerate(locations, 2):
        ws2.cell(row=r, column=1, value=loc)
        # B, C, D columns — EMPTY (task fills these)
        ws2.cell(row=r, column=2, value=None)
        ws2.cell(row=r, column=3, value=None)
        ws2.cell(row=r, column=4, value=None)

    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 26
    ws2.column_dimensions['C'].width = 16
    ws2.column_dimensions['D'].width = 22

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet HazmatRegister: 50 hazmat items, Hazard Class (D) empty, Storage Status (H) empty')
    print(f'  Sheet StorageSummary: 10 locations, B/C/D columns empty')

create_initial()
