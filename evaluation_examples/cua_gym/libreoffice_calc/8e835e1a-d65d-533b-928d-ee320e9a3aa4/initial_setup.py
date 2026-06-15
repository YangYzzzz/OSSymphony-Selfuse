"""
Initial Setup: Add data validation rule for text length on Schedule sheet
Task ID: calc_ggf_047
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_047'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Schedule ---
    ws = wb.active
    ws.title = 'Schedule'

    # Headers
    headers = ['Event ID', 'Event Name', 'Date', 'Location']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 30

    # 25 rows of event data (rows 2-26)
    # Some event names are intentionally too short or too long to motivate the validation rule
    events = [
        ['EVT-001', 'Q1 Revenue Strategy Kickoff Meeting', '2025-03-15', 'Grand Ballroom, Marriott Downtown'],
        ['EVT-002', 'AB', '2025-03-22', 'Conference Room 4B'],
        ['EVT-003', 'Spring Product Launch Celebration', '2025-04-01', 'Innovation Hub, Building 7'],
        ['EVT-004', 'Annual Employee Wellness Fair and Health Screening Event for All Departments Across Regional Offices', '2025-04-10', 'Fitness Center Atrium'],
        ['EVT-005', 'Customer Success Workshop', '2025-04-18', 'Training Center East'],
        ['EVT-006', 'X', '2025-04-25', 'Room 301'],
        ['EVT-007', 'Engineering All-Hands Update', '2025-05-02', 'Auditorium A'],
        ['EVT-008', 'Regional Sales Pipeline Review', '2025-05-09', 'Board Room 2'],
        ['EVT-009', 'New Hire Orientation Session', '2025-05-16', 'HR Training Lab'],
        ['EVT-010', 'Summer Intern Welcome Mixer', '2025-05-23', 'Rooftop Terrace'],
        ['EVT-011', 'QA', '2025-06-01', 'Testing Lab B'],
        ['EVT-012', 'Cross-Functional Team Building', '2025-06-08', 'Outdoor Pavilion'],
        ['EVT-013', 'Finance Quarter Close Review', '2025-06-15', 'Finance Dept, Room 510'],
        ['EVT-014', 'Marketing Campaign Brainstorm', '2025-06-22', 'Creative Studio'],
        ['EVT-015', 'IT Security Awareness Training', '2025-07-01', 'Virtual Meeting Room'],
        ['EVT-016', 'Mid-Year Performance Reviews for All Senior Management Staff Including Regional Directors and VP-Level Executives', '2025-07-10', 'Executive Suite'],
        ['EVT-017', 'Vendor Partnership Summit', '2025-07-18', 'Convention Center Hall C'],
        ['EVT-018', 'R&D Innovation Showcase', '2025-07-25', 'Demo Theater'],
        ['EVT-019', 'Compliance Training Update', '2025-08-02', 'Legal Conference Room'],
        ['EVT-020', 'H', '2025-08-10', 'Room 105'],
        ['EVT-021', 'Diversity and Inclusion Forum', '2025-08-18', 'Community Center'],
        ['EVT-022', 'Project Mercury Status Update', '2025-08-25', 'War Room Alpha'],
        ['EVT-023', 'Annual Company Picnic', '2025-09-05', 'Riverside Park'],
        ['EVT-024', 'Leadership Development Series', '2025-09-12', 'Executive Training Center'],
        ['EVT-025', 'End of Year Planning Session', '2025-09-20', 'Strategy Room'],
    ]

    for r, row_data in enumerate(events, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border

    # NO data validation on B2:B26 - that's the task

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
