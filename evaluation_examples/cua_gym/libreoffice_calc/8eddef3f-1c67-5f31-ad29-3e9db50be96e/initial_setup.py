"""
Initial Setup: VLOOKUP approximate match bug - lookup returns wrong price
Task ID: calc_tbl_041
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_041'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Products (lookup table, intentionally UNSORTED) ---
    ws_products = wb.active
    ws_products.title = "Products"

    headers = ["Product Name", "Description", "Price"]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")

    for col, h in enumerate(headers, 1):
        cell = ws_products.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align

    # Unsorted product data - this is the root cause of the VLOOKUP bug
    products = [
        ["Widget A", "Standard widget, blue finish", 9.99],
        ["Widget C", "Premium widget, chrome plated", 24.50],
        ["Widget B", "Mid-range widget, matte black", 14.99],
        ["Widget D", "Economy widget, plastic body", 5.49],
        ["Gadget Pro", "Electronic gadget with LCD display", 89.95],
        ["Widget E", "Heavy-duty widget, steel alloy", 32.00],
        ["Connector X", "Universal connector, brass fittings", 7.25],
        ["Widget F", "Miniature widget, precision cut", 18.75],
        ["Bracket Z", "L-shaped bracket, galvanized", 3.99],
        ["Widget G", "Custom widget, limited edition", 45.00],
        ["Gasket M", "Silicone gasket, heat resistant", 2.15],
        ["Widget H", "Oversized widget, industrial grade", 55.80],
    ]

    price_fmt = '$#,##0.00'
    for r, row_data in enumerate(products, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_products.cell(row=r, column=c, value=val)
            if c == 3:
                cell.number_format = price_fmt

    ws_products.column_dimensions["A"].width = 18
    ws_products.column_dimensions["B"].width = 38
    ws_products.column_dimensions["C"].width = 12

    # --- Sheet 2: Orders ---
    ws_orders = wb.create_sheet("Orders")

    order_headers = ["Order Item", "Quantity", "Unit", "Lookup Price"]
    for col, h in enumerate(order_headers, 1):
        cell = ws_orders.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align

    # Row 2: The problematic VLOOKUP with approximate match (1) on unsorted data
    ws_orders.cell(row=2, column=1, value="Widget B")
    ws_orders.cell(row=2, column=2, value=50)
    ws_orders.cell(row=2, column=3, value="pcs")
    # BUG: approximate match (1) on unsorted Products table returns wrong price
    ws_orders.cell(row=2, column=4, value='=VLOOKUP("Widget B",Products.A:C,3,1)')
    ws_orders["D2"].number_format = price_fmt

    # Additional order rows with static prices (not affected by the bug)
    more_orders = [
        ["Connector X", 100, "pcs", 7.25],
        ["Bracket Z", 200, "pcs", 3.99],
        ["Gadget Pro", 10, "pcs", 89.95],
        ["Gasket M", 500, "pcs", 2.15],
        ["Widget D", 75, "pcs", 5.49],
        ["Widget E", 30, "pcs", 32.00],
        ["Widget F", 60, "pcs", 18.75],
    ]
    for r, row_data in enumerate(more_orders, 3):
        for c, val in enumerate(row_data, 1):
            cell = ws_orders.cell(row=r, column=c, value=val)
            if c == 4:
                cell.number_format = price_fmt

    ws_orders.column_dimensions["A"].width = 18
    ws_orders.column_dimensions["B"].width = 12
    ws_orders.column_dimensions["C"].width = 10
    ws_orders.column_dimensions["D"].width = 14

    # Make Orders the active sheet (user sees the problem here)
    wb.active = wb.sheetnames.index("Orders")

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
