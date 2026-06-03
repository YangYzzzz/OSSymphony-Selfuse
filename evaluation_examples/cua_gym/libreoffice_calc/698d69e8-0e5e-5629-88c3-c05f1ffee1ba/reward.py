"""
Reward Script: Student Attendance Register
Task ID: calc_grs_031
Domain: libreoffice_calc
Scoring:
  C1 (0.20) - Structure: 25 students, 90 date cols, 4 summary cols
  C2 (0.15) - Data validation: list P,A,L,E on attendance range
  C3 (0.20) - Summary formulas: COUNTIF for Present/Absent/Late + Attendance %
  C4 (0.15) - Daily totals row with COUNTIF formulas
  C5 (0.10) - Conditional formatting on attendance data (P/A/L/E colors)
  C6 (0.10) - Conditional formatting on Attendance % (<75% flagged)
  C7 (0.10) - Freeze panes at B2
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_031'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Find the attendance sheet (may be named differently)
    ws = None
    for sn in wb.sheetnames:
        test_ws = wb[sn]
        # Look for a sheet that has student names and date-like headers
        if test_ws.max_row >= 20 and test_ws.max_column >= 50:
            ws = test_ws
            break
    if ws is None:
        # Fall back to first sheet
        ws = wb.worksheets[0]

    # ---------------------------------------------------------------
    # Component 1: Structure — 25 students, ~90 date cols, 4 summary cols (0.20 pts)
    # ---------------------------------------------------------------
    try:
        # Count student rows (rows 2..N where col A has non-empty text names)
        student_count = 0
        last_student_row = 1
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if val is not None and isinstance(val, str) and len(val.strip()) > 1:
                # Check it's a name, not a totals label
                lower_val = val.strip().lower()
                if 'total' in lower_val or 'count' in lower_val or 'daily' in lower_val:
                    continue
                student_count += 1
                last_student_row = r

        # Count date columns (columns with short date-like headers in row 1)
        date_col_count = 0
        first_date_col = None
        last_date_col = None
        for c in range(2, ws.max_column + 1):
            h = ws.cell(row=1, column=c).value
            if h is not None:
                h_str = str(h).strip()
                # Check if it looks like a date (MM/DD, or contains /)
                if re.match(r'^\d{1,2}/\d{1,2}', h_str):
                    date_col_count += 1
                    if first_date_col is None:
                        first_date_col = c
                    last_date_col = c

        # Check for summary column headers after date columns
        summary_headers_found = 0
        summary_keywords = ['present', 'absent', 'late', 'attendance', 'percentage', '%']
        summary_start_col = None
        if last_date_col:
            for c in range(last_date_col + 1, ws.max_column + 1):
                h = ws.cell(row=1, column=c).value
                if h is not None:
                    h_lower = str(h).strip().lower()
                    if any(kw in h_lower for kw in summary_keywords):
                        summary_headers_found += 1
                        if summary_start_col is None:
                            summary_start_col = c

        structure_ok = (student_count >= 24 and date_col_count >= 85 and summary_headers_found >= 3)
        if structure_ok:
            print(f"PASS: Component 1 — Structure: {student_count} students, {date_col_count} date cols, {summary_headers_found} summary cols (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 — students={student_count} (need>=24), date_cols={date_col_count} (need>=85), summary={summary_headers_found} (need>=3)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ---------------------------------------------------------------
    # Component 2: Data Validation — list with P,A,L,E (0.15 pts)
    # ---------------------------------------------------------------
    try:
        dvs = ws.data_validations.dataValidation
        matching_dvs = [
            dv for dv in dvs
            if dv.type == 'list' and dv.formula1
            and all(ch in dv.formula1.upper().replace(' ', '') for ch in ['P', 'A', 'L', 'E'])
        ]

        if len(matching_dvs) > 0:
            print(f"PASS: Component 2 — Data validation with P,A,L,E found (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 — No data validation with P,A,L,E list found. DVs: {len(dvs)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---------------------------------------------------------------
    # Component 3: Summary Formulas — COUNTIF for Present/Absent/Late + Attendance % (0.20 pts)
    # ---------------------------------------------------------------
    try:
        def check_summary_row(row_num):
            """Check if a row has COUNTIF formulas and attendance % formula in summary cols."""
            countifs = 0
            pct_ok = 0
            for c in range(summary_start_col, ws.max_column + 1):
                val = ws.cell(row=row_num, column=c).value
                if val and isinstance(val, str):
                    val_upper = val.upper().replace(' ', '')
                    if 'COUNTIF' in val_upper:
                        countifs += 1
                    header = str(ws.cell(row=1, column=c).value or '').lower()
                    if '%' in header or 'attendance' in header or 'percentage' in header:
                        if '/' in val or 'COUNTIF' in val_upper:
                            pct_ok += 1
            return countifs, pct_ok

        summary_formulas_ok = 0
        if summary_start_col:
            countif_found, pct_count = check_summary_row(2)
            if countif_found >= 3 and pct_count >= 1:
                summary_formulas_ok += 1

            countif_found_last, pct_count_last = check_summary_row(last_student_row)
            if countif_found_last >= 3 and pct_count_last >= 1:
                summary_formulas_ok += 1

            if summary_formulas_ok >= 2:
                print(f"PASS: Component 3 — Summary formulas verified for multiple rows: {countif_found} COUNTIFs + % formula (0.20 pts)")
                total_score += 0.20
            elif summary_formulas_ok >= 1:
                print(f"PARTIAL: Component 3 — Summary formulas verified for 1 row (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 3 — countif_found={countif_found}, pct_count={pct_count}")
        else:
            print(f"FAIL: Component 3 — No summary columns detected")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ---------------------------------------------------------------
    # Component 4: Daily Totals Row — COUNTIF formulas at bottom (0.15 pts)
    # ---------------------------------------------------------------
    try:
        # Find the totals row: a row after the last student with COUNTIF in date columns
        totals_row = None
        for r in range(last_student_row + 1, ws.max_row + 1):
            label = ws.cell(row=r, column=1).value
            if label is not None:
                label_str = str(label).lower()
                if 'total' in label_str or 'daily' in label_str or 'count' in label_str:
                    totals_row = r
                    break

        if totals_row is None:
            # Try checking the row right after last student
            for r in range(last_student_row + 1, min(last_student_row + 3, ws.max_row + 1)):
                val = ws.cell(row=r, column=2).value
                if val and isinstance(val, str) and 'COUNTIF' in val.upper():
                    totals_row = r
                    break

        if totals_row:
            # Check that multiple date columns in this row have COUNTIF formulas
            countif_count = 0
            sample_cols = [first_date_col, first_date_col + 10, first_date_col + 40, last_date_col] if first_date_col and last_date_col else []
            for c in sample_cols:
                if c and c <= ws.max_column:
                    val = ws.cell(row=totals_row, column=c).value
                    if val and isinstance(val, str) and 'COUNTIF' in val.upper():
                        countif_count += 1

            if countif_count >= 3:
                print(f"PASS: Component 4 — Daily totals row {totals_row} has COUNTIF formulas ({countif_count} sampled) (0.15 pts)")
                total_score += 0.15
            elif countif_count >= 1:
                print(f"PARTIAL: Component 4 — Daily totals row found but only {countif_count} COUNTIF formulas sampled (0.07 pts)")
                total_score += 0.07
            else:
                print(f"FAIL: Component 4 — Totals row {totals_row} found but no COUNTIF formulas in date columns")
        else:
            print(f"FAIL: Component 4 — No daily totals row found after student rows")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ---------------------------------------------------------------
    # Component 5: Conditional Formatting on attendance data — P/A/L/E colors (0.10 pts)
    # ---------------------------------------------------------------
    try:
        cf_rules = ws.conditional_formatting
        attendance_cf_count = 0
        target_values = set()

        for cf in cf_rules:
            cf_range = str(cf)
            # We want CF rules that cover the attendance data area (not just summary)
            for rule in cf.rules:
                if rule.type == 'cellIs' and hasattr(rule, 'formula') and rule.formula:
                    formula_str = str(rule.formula[0]).strip().strip('"').upper()
                    if formula_str in ('P', 'A', 'L', 'E'):
                        target_values.add(formula_str)
                        attendance_cf_count += 1

        if len(target_values) >= 4:
            print(f"PASS: Component 5 — Conditional formatting for {target_values} found (0.10 pts)")
            total_score += 0.10
        elif len(target_values) >= 2:
            print(f"PARTIAL: Component 5 — Conditional formatting for only {target_values} (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — Attendance CF rules found: {attendance_cf_count}, values: {target_values}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # ---------------------------------------------------------------
    # Component 6: CF on Attendance % column — <75% flagged (0.10 pts)
    # ---------------------------------------------------------------
    try:
        cf_rules = ws.conditional_formatting
        lt_75_count = 0
        for cf in cf_rules:
            for rule in cf.rules:
                if rule.type == 'cellIs' and hasattr(rule, 'formula') and rule.formula:
                    op = getattr(rule, 'operator', '')
                    if op == 'lessThan':
                        formula_val = str(rule.formula[0]).strip()
                        try:
                            fval = float(formula_val)
                            if abs(fval - 0.75) < 0.01 or abs(fval - 75) < 1:
                                lt_75_count += 1
                        except ValueError:
                            pass

        if lt_75_count > 0:
            print(f"PASS: Component 6 — Attendance % CF (<75% flagged) found (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 6 — No CF rule for <75% attendance found")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # ---------------------------------------------------------------
    # Component 7: Freeze Panes at B2 (row 1 + col A frozen) (0.10 pts)
    # ---------------------------------------------------------------
    try:
        fp = ws.freeze_panes
        if fp is not None and str(fp).upper() == 'B2':
            print(f"PASS: Component 7 — Freeze panes at B2 (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 7 — Freeze panes: {fp} (expected B2)")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
