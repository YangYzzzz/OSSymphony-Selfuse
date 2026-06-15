"""
Reward Script: Calculate shift differential pay
Task ID: calc_hr_shift_differential_pay_074
Domain: libreoffice_calc

Scoring rubric (total 1.0):
  Component 1: F2:F84 contain IFS formula for differential %  — 0.25 pts
  Component 2: G2:G84 contain adjusted rate formula           — 0.20 pts
  Component 3: H2:H84 contain total pay formula               — 0.20 pts
  Component 4: Number formats (F=0%, G=$#,##0.00, H=$#,##0.00)— 0.15 pts
  Component 5: Summary section (I/J headers + 4 SUMIFS rows)  — 0.20 pts
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_shift_differential_pay_074'

DATA_ROWS = range(2, 85)   # rows 2 through 84 inclusive (83 employees)
SUMMARY_SHIFT_TYPES = ['Day', 'Evening', 'Night', 'Weekend']


def normalize_formula(formula_str):
    """Strip leading '=' and remove all spaces for loose comparison."""
    if not isinstance(formula_str, str):
        return ''
    s = formula_str.strip()
    if s.startswith('='):
        s = s[1:]
    return s.replace(' ', '').upper()


def check_ifs_formula(formula_str, row):
    """
    Check that the formula matches the IFS differential pattern for a given row.
    Expected (case-insensitive, spaces stripped):
      =IFS(C{row}="Day",0,C{row}="Evening",0.10,C{row}="Night",0.15,C{row}="Weekend",0.20)
    We accept minor numeric variants for 0 (e.g. 0.00) and 0.10 etc.
    """
    if not isinstance(formula_str, str):
        return False
    normed = normalize_formula(formula_str)
    # Must start with IFS( and reference CX
    pattern = rf'^IFS\(C{row}="DAY"'
    return bool(re.match(pattern, normed, re.IGNORECASE))


def check_adjusted_rate_formula(formula_str, row):
    """G{row} = =E{row}*(1+F{row})"""
    if not isinstance(formula_str, str):
        return False
    normed = normalize_formula(formula_str)
    expected = f'E{row}*(1+F{row})'
    return normed == expected.upper()


def check_total_pay_formula(formula_str, row):
    """H{row} = =D{row}*G{row}"""
    if not isinstance(formula_str, str):
        return False
    normed = normalize_formula(formula_str)
    expected = f'D{row}*G{row}'
    return normed == expected.upper()


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Shift Pay' sheet must exist
    if 'Shift Pay' not in wb.sheetnames:
        print("CRITICAL: 'Shift Pay' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Shift Pay']

    # -------------------------------------------------------------------------
    # Component 1: F2:F84 contain IFS formula for differential % (0.25 pts)
    # Expected: =IFS(C{row}="Day",0,C{row}="Evening",0.10,C{row}="Night",0.15,C{row}="Weekend",0.20)
    # -------------------------------------------------------------------------
    try:
        f_pass = 0
        f_fail_examples = []
        for row in DATA_ROWS:
            val = ws.cell(row=row, column=6).value
            if check_ifs_formula(val, row):
                f_pass += 1
            else:
                f_fail_examples.append((row, val))

        total_data_rows = len(DATA_ROWS)
        f_fraction = f_pass / total_data_rows

        if f_fraction == 1.0:
            total_score += 0.25
            print(f"PASS: Component 1 — All {total_data_rows} F-column IFS formulas correct (0.25 pts)")
        elif f_fraction >= 0.8:
            partial = round(0.25 * f_fraction, 4)
            total_score += partial
            print(f"PARTIAL: Component 1 — {f_pass}/{total_data_rows} F-column IFS formulas correct "
                  f"({partial} pts). First failures: {f_fail_examples[:3]}")
        else:
            print(f"FAIL: Component 1 — Only {f_pass}/{total_data_rows} F-column IFS formulas correct. "
                  f"First failures: {f_fail_examples[:3]}")
    except Exception as e:
        print(f"ERROR: Component 1 (IFS formulas): {e}")

    # -------------------------------------------------------------------------
    # Component 2: G2:G84 contain adjusted rate formula =E{row}*(1+F{row}) (0.20 pts)
    # -------------------------------------------------------------------------
    try:
        g_pass = 0
        g_fail_examples = []
        for row in DATA_ROWS:
            val = ws.cell(row=row, column=7).value
            if check_adjusted_rate_formula(val, row):
                g_pass += 1
            else:
                g_fail_examples.append((row, val))

        g_fraction = g_pass / len(DATA_ROWS)

        if g_fraction == 1.0:
            total_score += 0.20
            print(f"PASS: Component 2 — All {len(DATA_ROWS)} G-column adjusted rate formulas correct (0.20 pts)")
        elif g_fraction >= 0.8:
            partial = round(0.20 * g_fraction, 4)
            total_score += partial
            print(f"PARTIAL: Component 2 — {g_pass}/{len(DATA_ROWS)} G-column formulas correct "
                  f"({partial} pts). First failures: {g_fail_examples[:3]}")
        else:
            print(f"FAIL: Component 2 — Only {g_pass}/{len(DATA_ROWS)} G-column formulas correct. "
                  f"First failures: {g_fail_examples[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 (adjusted rate formulas): {e}")

    # -------------------------------------------------------------------------
    # Component 3: H2:H84 contain total pay formula =D{row}*G{row} (0.20 pts)
    # -------------------------------------------------------------------------
    try:
        h_pass = 0
        h_fail_examples = []
        for row in DATA_ROWS:
            val = ws.cell(row=row, column=8).value
            if check_total_pay_formula(val, row):
                h_pass += 1
            else:
                h_fail_examples.append((row, val))

        h_fraction = h_pass / len(DATA_ROWS)

        if h_fraction == 1.0:
            total_score += 0.20
            print(f"PASS: Component 3 — All {len(DATA_ROWS)} H-column total pay formulas correct (0.20 pts)")
        elif h_fraction >= 0.8:
            partial = round(0.20 * h_fraction, 4)
            total_score += partial
            print(f"PARTIAL: Component 3 — {h_pass}/{len(DATA_ROWS)} H-column formulas correct "
                  f"({partial} pts). First failures: {h_fail_examples[:3]}")
        else:
            print(f"FAIL: Component 3 — Only {h_pass}/{len(DATA_ROWS)} H-column formulas correct. "
                  f"First failures: {h_fail_examples[:3]}")
    except Exception as e:
        print(f"ERROR: Component 3 (total pay formulas): {e}")

    # -------------------------------------------------------------------------
    # Component 4: Number formats — F: 0%, G: $#,##0.00, H: $#,##0.00 (0.15 pts)
    # Check a sample of cells (rows 2, 40, 84) per column
    # -------------------------------------------------------------------------
    try:
        format_checks = {
            'F (differential %)': (6, '0%'),
            'G (adjusted rate)': (7, '$#,##0.00'),
            'H (total pay)': (8, '$#,##0.00'),
        }
        sample_rows = [2, 40, 84]
        format_score = 0.0

        for label, (col, expected_fmt) in format_checks.items():
            col_pass = all(
                ws.cell(row=r, column=col).number_format == expected_fmt
                for r in sample_rows
            )
            if col_pass:
                format_score += 0.05
                print(f"PASS: Component 4 — {label} number format '{expected_fmt}' correct")
            else:
                actual_fmts = [ws.cell(row=r, column=col).number_format for r in sample_rows]
                print(f"FAIL: Component 4 — {label} format expected '{expected_fmt}', "
                      f"found in sample rows: {actual_fmts}")

        if format_score > 0:
            total_score += format_score
        if format_score == 0.15:
            print(f"PASS: Component 4 — All number formats correct (0.15 pts)")
    except Exception as e:
        print(f"ERROR: Component 4 (number formats): {e}")

    # -------------------------------------------------------------------------
    # Component 5: Summary section — I1/J1 headers + SUMIFS rows in I/J 87-90 (0.20 pts)
    # I1='Shift Type', J1='Total Differential Cost'
    # I87-I90 = shift types, J87-J90 = SUMIFS differential cost formulas
    # -------------------------------------------------------------------------
    try:
        summary_score = 0.0

        # Check headers in row 1 cols I (9) and J (10)
        i1 = ws.cell(row=1, column=9).value
        j1 = ws.cell(row=1, column=10).value
        headers_ok = (
            isinstance(i1, str) and 'shift' in i1.lower() and
            isinstance(j1, str) and 'differential' in j1.lower()
        )
        if headers_ok:
            summary_score += 0.05
            print(f"PASS: Component 5a — Summary headers correct: I1={repr(i1)}, J1={repr(j1)}")
        else:
            print(f"FAIL: Component 5a — Summary headers missing/wrong: I1={repr(i1)}, J1={repr(j1)}")

        # Check shift type labels in I87:I90
        shift_labels_found = []
        for row in range(87, 91):
            val = ws.cell(row=row, column=9).value
            if isinstance(val, str):
                shift_labels_found.append(val.strip())

        expected_shifts = set(SUMMARY_SHIFT_TYPES)
        found_shifts = set(shift_labels_found)
        if expected_shifts == found_shifts:
            summary_score += 0.05
            print(f"PASS: Component 5b — All 4 shift type labels found in I87:I90: {shift_labels_found}")
        else:
            missing = expected_shifts - found_shifts
            extra = found_shifts - expected_shifts
            print(f"FAIL: Component 5b — Shift labels mismatch. Missing: {missing}, Extra: {extra}, Found: {shift_labels_found}")

        # Check J87:J90 contain SUMIFS formulas (differential cost)
        j_formula_count = 0
        j_fail_examples = []
        for row in range(87, 91):
            val = ws.cell(row=row, column=10).value
            if isinstance(val, str) and 'SUMIFS' in val.upper():
                j_formula_count += 1
            else:
                j_fail_examples.append((row, val))

        if j_formula_count == 4:
            summary_score += 0.10
            print(f"PASS: Component 5c — All 4 SUMIFS differential cost formulas present in J87:J90 (0.10 pts)")
        elif j_formula_count > 0:
            partial = round(0.10 * j_formula_count / 4, 4)
            summary_score += partial
            print(f"PARTIAL: Component 5c — {j_formula_count}/4 SUMIFS formulas found ({partial} pts). "
                  f"Missing: {j_fail_examples}")
        else:
            print(f"FAIL: Component 5c — No SUMIFS formulas found in J87:J90. Found: {j_fail_examples}")

        if summary_score > 0:
            total_score += summary_score
        print(f"Component 5 subtotal: {summary_score}/0.20")
    except Exception as e:
        print(f"ERROR: Component 5 (summary section): {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
