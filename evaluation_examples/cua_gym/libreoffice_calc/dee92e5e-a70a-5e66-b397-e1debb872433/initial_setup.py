"""
Initial Setup: Shift differential pay calculation spreadsheet
Task ID: calc_hr_shift_differential_pay_074
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_shift_differential_pay_074'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Shift Pay'

    # Row 1: Headers (A-H)
    headers = ['Emp ID', 'Name', 'Shift Type', 'Hours Worked', 'Base Hourly Rate',
               'Differential %', 'Adjusted Rate', 'Total Pay']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Shift types distribution
    shift_types = ['Day', 'Evening', 'Night', 'Weekend']

    # Employee data - 83 employees (rows 2-84)
    employees = [
        ('EMP001', 'Sarah Chen', 'Evening', 40, 22.50),
        ('EMP002', 'Marcus Johnson', 'Night', 36, 25.00),
        ('EMP003', 'Priya Patel', 'Day', 40, 20.75),
        ('EMP004', 'James O\'Brien', 'Weekend', 24, 28.00),
        ('EMP005', 'Aisha Williams', 'Night', 40, 24.50),
        ('EMP006', 'David Kim', 'Day', 38, 21.00),
        ('EMP007', 'Fatima Al-Hassan', 'Evening', 40, 23.00),
        ('EMP008', 'Robert Martinez', 'Weekend', 16, 30.00),
        ('EMP009', 'Yuki Tanaka', 'Day', 40, 19.50),
        ('EMP010', 'Olivia Thompson', 'Night', 32, 26.75),
        ('EMP011', 'Carlos Rivera', 'Evening', 40, 22.00),
        ('EMP012', 'Mei-Ling Zhou', 'Day', 40, 20.25),
        ('EMP013', 'Kevin O\'Connor', 'Weekend', 24, 27.50),
        ('EMP014', 'Natasha Ivanova', 'Night', 40, 25.50),
        ('EMP015', 'Benjamin Park', 'Day', 36, 21.75),
        ('EMP016', 'Amara Osei', 'Evening', 40, 23.50),
        ('EMP017', 'Tyler Brooks', 'Night', 28, 24.00),
        ('EMP018', 'Isabela Costa', 'Day', 40, 20.00),
        ('EMP019', 'Ahmed Farouk', 'Weekend', 32, 29.00),
        ('EMP020', 'Chloe Anderson', 'Evening', 40, 22.75),
        ('EMP021', 'Dmitri Volkov', 'Day', 40, 21.25),
        ('EMP022', 'Grace Okonkwo', 'Night', 40, 26.00),
        ('EMP023', 'Lucas Silva', 'Evening', 36, 23.25),
        ('EMP024', 'Hannah Schmidt', 'Day', 40, 19.75),
        ('EMP025', 'Omar Abdullah', 'Weekend', 20, 31.00),
        ('EMP026', 'Valentina Cruz', 'Night', 32, 25.25),
        ('EMP027', 'Michael Taylor', 'Day', 40, 20.50),
        ('EMP028', 'Yolanda Brown', 'Evening', 40, 24.00),
        ('EMP029', 'Raj Sharma', 'Night', 36, 27.00),
        ('EMP030', 'Eleanor White', 'Day', 40, 21.50),
        ('EMP031', 'Jose Hernandez', 'Weekend', 24, 28.50),
        ('EMP032', 'Linh Nguyen', 'Evening', 40, 22.25),
        ('EMP033', 'Patrick Sullivan', 'Day', 38, 20.75),
        ('EMP034', 'Zara Ahmed', 'Night', 40, 26.50),
        ('EMP035', 'Samuel Green', 'Day', 40, 21.00),
        ('EMP036', 'Keiko Yamamoto', 'Evening', 32, 23.75),
        ('EMP037', 'Andre Dumont', 'Weekend', 16, 32.00),
        ('EMP038', 'Nadia Popescu', 'Night', 40, 25.75),
        ('EMP039', 'Christopher Lee', 'Day', 40, 22.00),
        ('EMP040', 'Abebe Girma', 'Evening', 40, 24.50),
        ('EMP041', 'Svetlana Kozlov', 'Day', 36, 20.25),
        ('EMP042', 'Ethan Murphy', 'Night', 28, 27.25),
        ('EMP043', 'Layla Hassan', 'Weekend', 32, 29.50),
        ('EMP044', 'Nikolai Petrov', 'Day', 40, 21.75),
        ('EMP045', 'Simone Leblanc', 'Evening', 40, 23.00),
        ('EMP046', 'Arjun Mehta', 'Night', 40, 26.25),
        ('EMP047', 'Florence Dubois', 'Day', 40, 20.00),
        ('EMP048', 'Hassan Mukhtar', 'Evening', 36, 22.50),
        ('EMP049', 'Ingrid Larsson', 'Weekend', 24, 30.50),
        ('EMP050', 'Jamal Washington', 'Night', 32, 25.00),
        ('EMP051', 'Maria Rossi', 'Day', 40, 21.25),
        ('EMP052', 'Thomas Fischer', 'Evening', 40, 24.25),
        ('EMP053', 'Adia Kamara', 'Night', 36, 27.50),
        ('EMP054', 'Peng Liu', 'Day', 40, 20.75),
        ('EMP055', 'Sofia Morales', 'Weekend', 20, 31.50),
        ('EMP056', 'Viktor Novak', 'Evening', 40, 23.50),
        ('EMP057', 'Rashida Jones', 'Day', 38, 22.25),
        ('EMP058', 'Aleksei Smirnov', 'Night', 40, 26.75),
        ('EMP059', 'Carmen Gonzalez', 'Day', 40, 21.00),
        ('EMP060', 'Tobias Richter', 'Evening', 32, 23.00),
        ('EMP061', 'Salma Youssef', 'Weekend', 24, 28.75),
        ('EMP062', 'Nathan Davis', 'Night', 40, 25.50),
        ('EMP063', 'Amira Mansour', 'Day', 40, 20.50),
        ('EMP064', 'Brendan O\'Neill', 'Evening', 40, 24.75),
        ('EMP065', 'Ayasha Whitehorse', 'Night', 28, 27.75),
        ('EMP066', 'Luca Ferrari', 'Day', 40, 21.50),
        ('EMP067', 'Thanh Tran', 'Weekend', 16, 33.00),
        ('EMP068', 'Miriam Goldstein', 'Evening', 40, 22.75),
        ('EMP069', 'Santiago Reyes', 'Day', 36, 20.25),
        ('EMP070', 'Chioma Eze', 'Night', 40, 26.00),
        ('EMP071', 'Pawel Kowalski', 'Day', 40, 21.75),
        ('EMP072', 'Rania Khalil', 'Evening', 40, 23.25),
        ('EMP073', 'Douglas Campbell', 'Weekend', 32, 30.00),
        ('EMP074', 'Soren Nielsen', 'Night', 36, 25.25),
        ('EMP075', 'Blessing Obi', 'Day', 40, 20.75),
        ('EMP076', 'Mikhail Orlov', 'Evening', 28, 24.00),
        ('EMP077', 'Tanya Robson', 'Night', 40, 27.00),
        ('EMP078', 'Fernanda Lima', 'Day', 40, 22.00),
        ('EMP079', 'Hiro Matsuda', 'Weekend', 20, 32.50),
        ('EMP080', 'Nkechi Okafor', 'Evening', 40, 23.75),
        ('EMP081', 'Ivan Petrov', 'Day', 38, 21.25),
        ('EMP082', 'Catalina Torres', 'Night', 40, 26.50),
        ('EMP083', 'Dae-Jung Kim', 'Day', 40, 20.00),
    ]

    for r, (emp_id, name, shift, hours, rate) in enumerate(employees, 2):
        ws.cell(row=r, column=1, value=emp_id)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=shift)
        ws.cell(row=r, column=4, value=hours)
        ws.cell(row=r, column=5, value=rate)
        # Columns F, G, H are intentionally left empty

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Shift Pay')
    print(f'  Rows 2-84 populated with shift worker data (83 employees)')
    print(f'  Columns A-E filled; F, G, H are empty')


create_initial()
