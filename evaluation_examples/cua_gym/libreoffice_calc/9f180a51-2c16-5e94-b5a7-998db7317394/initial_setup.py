"""
Initial Setup: HR Recruitment Pipeline Tracker
Task ID: calc_grs_051
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_051'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Candidates ---
    ws = wb.active
    ws.title = "Candidates"

    headers = [
        "Candidate Name", "Position Applied", "Department", "Date Applied",
        "Resume Source", "Current Stage", "Stage Date", "Hiring Manager",
        "Interviewer(s)", "Score", "Notes"
    ]

    # Style headers
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 20 candidates data - 5 open positions across departments
    candidates = [
        ["Sarah Chen", "Senior Software Engineer", "Engineering", "2025-11-05", "LinkedIn", "Final Interview", "2025-12-18", "David Park", "Lisa Wang, Tom Reed", 8, "Strong system design skills"],
        ["Marcus Johnson", "Senior Software Engineer", "Engineering", "2025-11-12", "Indeed", "Interview Round 2", "2025-12-10", "David Park", "Lisa Wang", 7, "Good coding, needs design round"],
        ["Priya Sharma", "Marketing Manager", "Marketing", "2025-10-28", "Referral", "Offer", "2025-12-20", "Rachel Kim", "Jake Foster, Nina Patel", 9, "Exceptional campaign portfolio"],
        ["James O'Brien", "Data Analyst", "Analytics", "2025-11-15", "Career Site", "Phone Screen", "2025-11-22", "Michelle Torres", "N/A", "", "Scheduled for next Tuesday"],
        ["Aisha Patel", "Senior Software Engineer", "Engineering", "2025-11-01", "Agency", "Hired", "2025-12-15", "David Park", "Lisa Wang, Tom Reed, Sam Liu", 9, "Started Jan 6th"],
        ["Carlos Rivera", "Product Designer", "Design", "2025-11-20", "LinkedIn", "Interview Round 1", "2025-12-05", "Emily Zhao", "Chris Martin", 6, "Portfolio review pending"],
        ["Emma Thompson", "Marketing Manager", "Marketing", "2025-11-08", "LinkedIn", "Rejected", "2025-12-01", "Rachel Kim", "Jake Foster", 4, "Insufficient B2B experience"],
        ["Wei Zhang", "Data Analyst", "Analytics", "2025-11-18", "Referral", "Interview Round 1", "2025-12-08", "Michelle Torres", "Omar Hassan", 7, "Strong SQL skills demonstrated"],
        ["Olivia Martinez", "Senior Software Engineer", "Engineering", "2025-11-25", "Career Site", "Application", "2025-11-25", "David Park", "N/A", "", "Resume under review"],
        ["Ryan Kowalski", "Product Designer", "Design", "2025-10-30", "Indeed", "Interview Round 2", "2025-12-12", "Emily Zhao", "Chris Martin, Jess Lee", 8, "Excellent Figma prototypes"],
        ["Fatima Al-Rashid", "Marketing Manager", "Marketing", "2025-11-22", "Agency", "Phone Screen", "2025-12-02", "Rachel Kim", "N/A", "", "Agency highly recommends"],
        ["Daniel Nguyen", "Data Analyst", "Analytics", "2025-11-10", "LinkedIn", "Final Interview", "2025-12-16", "Michelle Torres", "Omar Hassan, Yuki Tanaka", 8, "Python and Tableau expert"],
        ["Hannah Brooks", "Senior Software Engineer", "Engineering", "2025-12-01", "Referral", "Application", "2025-12-01", "David Park", "N/A", "", "Referred by Aisha Patel"],
        ["Tomasz Lewandowski", "Product Designer", "Design", "2025-11-14", "LinkedIn", "Offer", "2025-12-19", "Emily Zhao", "Chris Martin, Jess Lee, Emily Zhao", 9, "Outstanding UX research approach"],
        ["Grace Kim", "Data Analyst", "Analytics", "2025-11-28", "Career Site", "Application", "2025-11-28", "Michelle Torres", "N/A", "", "MS in Statistics from Stanford"],
        ["Alex Volkov", "Senior Software Engineer", "Engineering", "2025-11-08", "Indeed", "Interview Round 1", "2025-12-04", "David Park", "Tom Reed", 6, "Backend focus, limited frontend"],
        ["Jasmine Washington", "Marketing Manager", "Marketing", "2025-11-16", "Referral", "Interview Round 2", "2025-12-14", "Rachel Kim", "Jake Foster, Nina Patel", 7, "Creative digital strategy"],
        ["Liam O'Connor", "Product Designer", "Design", "2025-12-02", "Career Site", "Phone Screen", "2025-12-09", "Emily Zhao", "N/A", "", "Interesting startup background"],
        ["Sofia Andersson", "Data Analyst", "Analytics", "2025-11-06", "Agency", "Hired", "2025-12-10", "Michelle Torres", "Omar Hassan, Yuki Tanaka", 8, "Joined analytics team Dec 16"],
        ["Kwame Asante", "Senior Software Engineer", "Engineering", "2025-11-19", "LinkedIn", "Interview Round 2", "2025-12-11", "David Park", "Lisa Wang, Sam Liu", 7, "Solid distributed systems background"],
    ]

    for r, row_data in enumerate(candidates, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 4 or c == 7:  # Date columns
                cell.number_format = 'yyyy-mm-dd'
            if c == 10 and val != "":  # Score column
                cell.alignment = Alignment(horizontal="center")

    # Data Validation - Resume Source dropdown
    dv_source = DataValidation(
        type="list",
        formula1='"LinkedIn,Indeed,Referral,Career Site,Agency"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_source.error = "Please select a valid resume source"
    dv_source.errorTitle = "Invalid Source"
    dv_source.add("E2:E100")
    ws.add_data_validation(dv_source)

    # Data Validation - Current Stage dropdown
    dv_stage = DataValidation(
        type="list",
        formula1='"Application,Phone Screen,Interview Round 1,Interview Round 2,Final Interview,Offer,Hired,Rejected"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_stage.error = "Please select a valid stage"
    dv_stage.errorTitle = "Invalid Stage"
    dv_stage.add("F2:F100")
    ws.add_data_validation(dv_stage)

    # Set column widths
    col_widths = {
        'A': 22, 'B': 26, 'C': 14, 'D': 14,
        'E': 16, 'F': 20, 'G': 14, 'H': 18,
        'I': 28, 'J': 8, 'K': 32
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Set row 1 height
    ws.row_dimensions[1].height = 30

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
