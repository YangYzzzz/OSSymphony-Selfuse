"""
Reward Script: HR Recruitment Pipeline Tracker
Task ID: calc_grs_051
Domain: libreoffice_calc
Scoring:
  Component 1: Conditional formatting on Current Stage column (0.25 pts)
  Component 2: Pipeline Summary sheet with stage counts (0.30 pts)
  Component 3: Time-to-Hire Analysis sheet (0.20 pts)
  Component 4: Source Effectiveness sheet (0.25 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_051'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    sheet_names = wb.sheetnames

    # =========================================================================
    # Component 1: Conditional formatting on Current Stage column (0.25 pts)
    # The task requires color-coded conditional formatting on the Current Stage
    # column (F). The initial file has NO conditional formatting rules.
    # Golden has 8 cellIs rules on F2:F21 for each stage value.
    # =========================================================================
    try:
        ws_cand = wb['Candidates']
        cf_rules = list(ws_cand.conditional_formatting)
        # Count rules that apply to column F (the Current Stage column)
        stage_cf_count = 0
        for cf in cf_rules:
            cf_range = str(cf)
            # Check if the conditional formatting range includes column F
            if 'F' in cf_range:
                for rule in cf.rules:
                    if rule.type == 'cellIs':
                        stage_cf_count += 1

        if stage_cf_count >= 6:
            # At least 6 distinct stage-based conditional formatting rules
            print(f"PASS: Component 1 — Found {stage_cf_count} cellIs conditional formatting rules on column F (0.25 pts)")
            total_score += 0.25
        elif stage_cf_count >= 3:
            # Partial credit for some rules
            print(f"PARTIAL: Component 1 — Found {stage_cf_count} cellIs rules (need >= 6 for full credit) (0.12 pts)")
            total_score += 0.12
        else:
            print(f"FAIL: Component 1 — Found {stage_cf_count} cellIs conditional formatting rules on column F, expected >= 6")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Pipeline Summary sheet with stage counts (0.30 pts)
    # The task asks for a "pipeline funnel summary showing count of candidates
    # at each stage". The initial file has no such sheet.
    # Golden has a 'Pipeline Summary' sheet with Stage/Count/% columns and
    # rows for each of the 8 stages.
    # =========================================================================
    try:
        # Check that a Pipeline Summary sheet (or similar name) exists
        pipeline_sheet = None
        for name in sheet_names:
            if 'pipeline' in name.lower() or 'funnel' in name.lower() or 'summary' in name.lower():
                pipeline_sheet = wb[name]
                break

        if pipeline_sheet is None:
            print("FAIL: Component 2 — No Pipeline Summary / Funnel sheet found")
        else:
            # Check for stage names and counts
            # Collect all cell values to find stage references
            all_values = []
            for r in range(1, pipeline_sheet.max_row + 1):
                for c in range(1, pipeline_sheet.max_column + 1):
                    v = pipeline_sheet.cell(row=r, column=c).value
                    if v is not None:
                        all_values.append(str(v))

            all_text = ' '.join(all_values)
            # The 8 stages from the task
            expected_stages = [
                'Application', 'Phone Screen', 'Interview Round 1',
                'Interview Round 2', 'Final Interview', 'Offer', 'Hired', 'Rejected'
            ]
            found_stages = [s for s in expected_stages if s in all_text]

            # Check for numeric counts (should have numbers representing candidate counts)
            has_counts = False
            for r in range(1, pipeline_sheet.max_row + 1):
                for c in range(1, pipeline_sheet.max_column + 1):
                    v = pipeline_sheet.cell(row=r, column=c).value
                    if isinstance(v, (int, float)) and 0 < v <= 20:
                        has_counts = True
                        break
                if has_counts:
                    break

            if len(found_stages) >= 6 and has_counts:
                print(f"PASS: Component 2 — Pipeline Summary sheet found with {len(found_stages)}/8 stages and numeric counts (0.30 pts)")
                total_score += 0.30
            elif len(found_stages) >= 3 and has_counts:
                print(f"PARTIAL: Component 2 — Pipeline Summary found with {len(found_stages)}/8 stages (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 2 — Pipeline Summary sheet found but missing stages ({len(found_stages)}/8) or counts (has_counts={has_counts})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Time-to-Hire Analysis sheet (0.20 pts)
    # The task asks for "time-to-hire analysis showing days between Application
    # and Hire date". Initial file has no such sheet.
    # Golden has 'Time-to-Hire Analysis' with hired candidates and days calc.
    # =========================================================================
    try:
        tth_sheet = None
        for name in sheet_names:
            if 'time' in name.lower() and 'hire' in name.lower():
                tth_sheet = wb[name]
                break
            elif 'hire' in name.lower() and ('analysis' in name.lower() or 'days' in name.lower()):
                tth_sheet = wb[name]
                break

        if tth_sheet is None:
            print("FAIL: Component 3 — No Time-to-Hire Analysis sheet found")
        else:
            # Look for hired candidate names and days-to-hire values
            all_values = []
            has_days_value = False
            has_hired_candidate = False
            # Known hired candidates from the data: Aisha Patel, Sofia Andersson
            hired_names = ['Aisha', 'Sofia', 'Patel', 'Andersson']

            for r in range(1, tth_sheet.max_row + 1):
                for c in range(1, tth_sheet.max_column + 1):
                    v = tth_sheet.cell(row=r, column=c).value
                    if v is not None:
                        sv = str(v)
                        all_values.append(sv)
                        for name in hired_names:
                            if name in sv:
                                has_hired_candidate = True
                        # Days to hire should be a positive integer (reasonable range 1-365)
                        if isinstance(v, (int, float)) and 1 <= v <= 365:
                            has_days_value = True

            all_text = ' '.join(all_values)
            has_days_header = 'days' in all_text.lower() or 'time' in all_text.lower()

            if has_hired_candidate and has_days_value:
                print(f"PASS: Component 3 — Time-to-Hire Analysis found with hired candidates and days values (0.20 pts)")
                total_score += 0.20
            elif has_hired_candidate or has_days_value:
                print(f"PARTIAL: Component 3 — Sheet found but incomplete (hired_candidate={has_hired_candidate}, days={has_days_value}) (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 3 — Time-to-Hire sheet found but no hired candidates or days values")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: Source Effectiveness sheet (0.25 pts)
    # The task asks for "source effectiveness table showing hire rates by
    # resume source". Initial file has no such sheet.
    # Golden has 'Source Effectiveness' with source names, application counts,
    # hires, and hire rates.
    # =========================================================================
    try:
        src_sheet = None
        for name in sheet_names:
            if 'source' in name.lower() or 'effectiveness' in name.lower():
                src_sheet = wb[name]
                break

        if src_sheet is None:
            print("FAIL: Component 4 — No Source Effectiveness sheet found")
        else:
            # Check for resume source names and hire rate data
            all_values = []
            expected_sources = ['LinkedIn', 'Indeed', 'Referral', 'Career Site', 'Agency']
            found_sources = []
            has_rate_or_count = False

            for r in range(1, src_sheet.max_row + 1):
                for c in range(1, src_sheet.max_column + 1):
                    v = src_sheet.cell(row=r, column=c).value
                    if v is not None:
                        sv = str(v)
                        all_values.append(sv)
                        for src in expected_sources:
                            if src.lower() in sv.lower() and src not in found_sources:
                                found_sources.append(src)
                        # Check for numeric values that could be counts or rates
                        if isinstance(v, (int, float)) and v >= 0:
                            has_rate_or_count = True

            all_text = ' '.join(all_values)
            has_hire_ref = 'hire' in all_text.lower() or 'rate' in all_text.lower()

            if len(found_sources) >= 4 and has_rate_or_count and has_hire_ref:
                print(f"PASS: Component 4 — Source Effectiveness found with {len(found_sources)}/5 sources and hire rate data (0.25 pts)")
                total_score += 0.25
            elif len(found_sources) >= 2 and has_rate_or_count:
                print(f"PARTIAL: Component 4 — Source Effectiveness found with {len(found_sources)}/5 sources (0.12 pts)")
                total_score += 0.12
            else:
                print(f"FAIL: Component 4 — Source Effectiveness sheet found but missing sources ({len(found_sources)}/5) or rate data")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
