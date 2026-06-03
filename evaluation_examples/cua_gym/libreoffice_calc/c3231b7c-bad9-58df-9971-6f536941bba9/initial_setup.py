"""
Initial Setup: Highlight overdue and upcoming-due dates with conditional formatting
Task ID: calc_gcv_025
Domain: libreoffice_calc

Creates a Task_Schedule spreadsheet with 34 tasks spanning 2025-12-01 to 2026-06-30.
No conditional formatting applied - that is the agent's job.
"""

import os
import shlex
import subprocess
import time
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_025'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Task_Schedule"

    # --- Headers ---
    headers = ["Task ID", "Task Name", "Owner", "Priority", "Start Date", "Due Date"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Column widths ---
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14

    # --- Task Data (34 rows) ---
    # Dates span from 2025-12-01 to 2026-06-30
    owners = [
        "Sarah Chen", "Marcus Johnson", "Elena Rodriguez", "David Kim",
        "Priya Patel", "James Wilson", "Aisha Okafor", "Robert Taylor",
        "Mei-Lin Wu", "Carlos Fernandez", "Anna Kowalski", "Thomas Brown",
    ]
    priorities = ["High", "Medium", "Low", "Critical"]

    tasks = [
        ("TSK-001", "Migrate legacy database to cloud infrastructure", "Sarah Chen", "Critical", date(2025, 10, 15), date(2025, 12, 1)),
        ("TSK-002", "Design new customer onboarding workflow", "Marcus Johnson", "High", date(2025, 11, 1), date(2025, 12, 15)),
        ("TSK-003", "Implement two-factor authentication module", "Elena Rodriguez", "Critical", date(2025, 11, 10), date(2025, 12, 28)),
        ("TSK-004", "Create quarterly financial report template", "David Kim", "Medium", date(2025, 11, 20), date(2026, 1, 5)),
        ("TSK-005", "Optimize search indexing pipeline", "Priya Patel", "High", date(2025, 12, 1), date(2026, 1, 15)),
        ("TSK-006", "Redesign mobile app navigation menu", "James Wilson", "Medium", date(2025, 12, 5), date(2026, 1, 25)),
        ("TSK-007", "Set up automated regression test suite", "Aisha Okafor", "High", date(2025, 12, 10), date(2026, 2, 1)),
        ("TSK-008", "Develop API rate limiting middleware", "Robert Taylor", "Critical", date(2025, 12, 15), date(2026, 2, 10)),
        ("TSK-009", "Audit third-party vendor security compliance", "Mei-Lin Wu", "High", date(2026, 1, 2), date(2026, 2, 20)),
        ("TSK-010", "Build real-time analytics dashboard", "Carlos Fernandez", "Medium", date(2026, 1, 5), date(2026, 2, 28)),
        ("TSK-011", "Refactor payment processing service", "Anna Kowalski", "Critical", date(2026, 1, 10), date(2026, 3, 1)),
        ("TSK-012", "Write user documentation for v3.0 release", "Thomas Brown", "Low", date(2026, 1, 15), date(2026, 3, 10)),
        ("TSK-013", "Configure CI/CD pipeline for microservices", "Sarah Chen", "High", date(2026, 1, 20), date(2026, 3, 15)),
        ("TSK-014", "Implement server-side caching strategy", "Marcus Johnson", "Medium", date(2026, 2, 1), date(2026, 3, 20)),
        ("TSK-015", "Perform load testing on checkout flow", "Elena Rodriguez", "High", date(2026, 2, 5), date(2026, 3, 25)),
        ("TSK-016", "Design employee performance review portal", "David Kim", "Medium", date(2026, 2, 10), date(2026, 3, 28)),
        ("TSK-017", "Integrate Slack notification webhooks", "Priya Patel", "Low", date(2026, 2, 15), date(2026, 3, 30)),
        ("TSK-018", "Deploy containerized staging environment", "James Wilson", "High", date(2026, 2, 20), date(2026, 3, 31)),
        ("TSK-019", "Create data backup and recovery procedures", "Aisha Okafor", "Critical", date(2026, 3, 1), date(2026, 4, 1)),
        ("TSK-020", "Upgrade frontend framework to React 19", "Robert Taylor", "Medium", date(2026, 3, 3), date(2026, 4, 2)),
        ("TSK-021", "Implement role-based access control updates", "Mei-Lin Wu", "High", date(2026, 3, 5), date(2026, 4, 3)),
        ("TSK-022", "Build automated invoice generation system", "Carlos Fernandez", "Medium", date(2026, 3, 8), date(2026, 4, 5)),
        ("TSK-023", "Conduct penetration testing on public APIs", "Anna Kowalski", "Critical", date(2026, 3, 10), date(2026, 4, 8)),
        ("TSK-024", "Develop customer feedback analytics module", "Thomas Brown", "Low", date(2026, 3, 12), date(2026, 4, 10)),
        ("TSK-025", "Optimize database query performance", "Sarah Chen", "High", date(2026, 3, 15), date(2026, 4, 15)),
        ("TSK-026", "Set up monitoring and alerting with Grafana", "Marcus Johnson", "Medium", date(2026, 3, 20), date(2026, 4, 20)),
        ("TSK-027", "Migrate email service to SendGrid platform", "Elena Rodriguez", "Low", date(2026, 3, 25), date(2026, 4, 30)),
        ("TSK-028", "Implement GraphQL subscriptions endpoint", "David Kim", "High", date(2026, 4, 1), date(2026, 5, 10)),
        ("TSK-029", "Design multi-tenant architecture proposal", "Priya Patel", "Critical", date(2026, 4, 5), date(2026, 5, 20)),
        ("TSK-030", "Build internal knowledge base search tool", "James Wilson", "Medium", date(2026, 4, 10), date(2026, 5, 30)),
        ("TSK-031", "Automate compliance report generation", "Aisha Okafor", "High", date(2026, 4, 15), date(2026, 6, 5)),
        ("TSK-032", "Develop mobile push notification service", "Robert Taylor", "Medium", date(2026, 4, 20), date(2026, 6, 15)),
        ("TSK-033", "Implement data anonymization for GDPR", "Mei-Lin Wu", "Critical", date(2026, 5, 1), date(2026, 6, 20)),
        ("TSK-034", "Create end-of-year project retrospective", "Carlos Fernandez", "Low", date(2026, 5, 15), date(2026, 6, 30)),
    ]

    date_format = 'YYYY-MM-DD'
    for r, (tid, name, owner, priority, start, due) in enumerate(tasks, 2):
        ws.cell(row=r, column=1, value=tid)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=owner)
        ws.cell(row=r, column=4, value=priority)

        start_cell = ws.cell(row=r, column=5, value=start)
        start_cell.number_format = date_format

        due_cell = ws.cell(row=r, column=6, value=due)
        due_cell.number_format = date_format

        # Light row styling for readability
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = thin_border

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
