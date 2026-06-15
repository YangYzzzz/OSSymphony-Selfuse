"""
Reward Script: Conditional formatting for due dates in Task_Schedule
Task ID: calc_gcv_025
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20): Conditional formatting exists on range F2:F35
  Component 2 (0.30): Yellow (#FFFF00) rule for upcoming dates (within 7 days)
  Component 3 (0.30): Orange (#FFA500) rule for past-due dates
  Component 4 (0.20): Correct rule priority (yellow=higher priority than orange)
"""

import os

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_025'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify conditional formatting rules on F2:F35 for date highlighting.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Precondition: sheet must be Task_Schedule with data
    if ws.title != "Task_Schedule":
        # Try to find it
        if "Task_Schedule" in wb.sheetnames:
            ws = wb["Task_Schedule"]
        else:
            print(f"FAIL: No 'Task_Schedule' sheet found. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0

    # Collect all conditional formatting rules that cover F2:F35
    cf_rules_on_f = []
    for cf in ws.conditional_formatting:
        range_str = str(cf).replace("<ConditionalFormatting ", "").replace(">", "").strip()
        # Check if the range covers F2:F35 (could be exact or superset)
        if "F2" in range_str and "F35" in range_str:
            for rule in cf.rules:
                cf_rules_on_f.append(rule)

    # Component 1: Conditional formatting exists on F2:F35 range (0.20 points)
    try:
        if len(cf_rules_on_f) >= 2:
            print(f"PASS: Component 1 -- Found {len(cf_rules_on_f)} conditional formatting rules on F2:F35 (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 -- Expected >= 2 CF rules on F2:F35, found {len(cf_rules_on_f)}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Yellow (#FFFF00) rule for upcoming dates within 7 days (0.30 points)
    yellow_rule_found = False
    try:
        for rule in cf_rules_on_f:
            formulas = [f.upper().replace(" ", "") for f in (rule.formula or [])]
            # Check for formula containing AND, TODAY, TODAY()+7 pattern
            has_upcoming_formula = False
            for f in formulas:
                if "TODAY()" in f and "+7" in f and "AND(" in f:
                    has_upcoming_formula = True
                    break
                # Also accept TODAY()+7 variants
                if "TODAY()+7" in f and ">=" in f:
                    has_upcoming_formula = True
                    break

            if has_upcoming_formula:
                # Check fill color is yellow FFFFFF00 or FFFF00
                if rule.dxf and rule.dxf.fill and rule.dxf.fill.fgColor:
                    color_rgb = rule.dxf.fill.fgColor.rgb
                    if color_rgb in ("FFFFFF00", "00FFFF00"):
                        yellow_rule_found = True
                        print(f"PASS: Component 2 -- Yellow rule found: formula={formulas}, fill={color_rgb} (0.30 pts)")
                        total_score += 0.30
                    else:
                        print(f"FAIL: Component 2 -- Upcoming-dates rule found but fill color is {color_rgb}, expected FFFFFF00")
                else:
                    print(f"FAIL: Component 2 -- Upcoming-dates rule found but no fill color defined")

        if not yellow_rule_found:
            # Check if we printed a FAIL already
            has_any_upcoming = any(
                any("TODAY()" in f.upper() and "+7" in f.upper() for f in (r.formula or []))
                for r in cf_rules_on_f
            )
            if not has_any_upcoming:
                print(f"FAIL: Component 2 -- No rule with upcoming-dates formula (AND, TODAY, +7) found")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Orange (#FFA500) rule for past-due dates (0.30 points)
    orange_rule_found = False
    try:
        for rule in cf_rules_on_f:
            formulas = [f.upper().replace(" ", "") for f in (rule.formula or [])]
            # Check for formula with F2<TODAY() pattern (past due)
            has_pastdue_formula = False
            for f in formulas:
                if "<TODAY()" in f and "AND(" not in f:
                    has_pastdue_formula = True
                    break
                # Also accept variants like F2<TODAY()
                if "F2<TODAY()" in f.replace(" ", "").upper():
                    has_pastdue_formula = True
                    break

            if has_pastdue_formula:
                if rule.dxf and rule.dxf.fill and rule.dxf.fill.fgColor:
                    color_rgb = rule.dxf.fill.fgColor.rgb
                    if color_rgb in ("FFFFA500", "00FFA500"):
                        orange_rule_found = True
                        print(f"PASS: Component 3 -- Orange rule found: formula={formulas}, fill={color_rgb} (0.30 pts)")
                        total_score += 0.30
                    else:
                        print(f"FAIL: Component 3 -- Past-due rule found but fill color is {color_rgb}, expected FFFFA500")
                else:
                    print(f"FAIL: Component 3 -- Past-due rule found but no fill color defined")

        if not orange_rule_found:
            has_any_pastdue = any(
                any("<TODAY()" in f.upper().replace(" ", "") for f in (r.formula or []))
                for r in cf_rules_on_f
            )
            if not has_any_pastdue:
                print(f"FAIL: Component 3 -- No rule with past-due formula (<TODAY()) found")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Correct priority ordering (0.20 points)
    # Yellow (upcoming) should have higher priority (lower number) than orange (past-due)
    try:
        if yellow_rule_found and orange_rule_found:
            yellow_priority = None
            orange_priority = None
            for rule in cf_rules_on_f:
                formulas = [f.upper().replace(" ", "") for f in (rule.formula or [])]
                for f in formulas:
                    if "TODAY()" in f and "+7" in f and "AND(" in f:
                        yellow_priority = rule.priority
                    if "<TODAY()" in f and "AND(" not in f:
                        orange_priority = rule.priority

            if yellow_priority is not None and orange_priority is not None:
                if yellow_priority < orange_priority:
                    print(f"PASS: Component 4 -- Yellow priority ({yellow_priority}) < Orange priority ({orange_priority}), correct order (0.20 pts)")
                    total_score += 0.20
                else:
                    print(f"FAIL: Component 4 -- Yellow priority ({yellow_priority}) should be < Orange priority ({orange_priority})")
            else:
                print(f"FAIL: Component 4 -- Could not determine priorities: yellow={yellow_priority}, orange={orange_priority}")
        else:
            print(f"FAIL: Component 4 -- Cannot check priority without both rules present")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
