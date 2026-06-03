"""
Initial Setup: 2D Lookup using INDEX and MATCH for shipping cost
Task ID: calc_fma_index_match_2d_005
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_fma_index_match_2d_005'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Shipping'

    # --- Rate matrix header row (row 1) ---
    ws['A1'] = None       # blank top-left corner
    ws['B1'] = 'Zone1'
    ws['C1'] = 'Zone2'
    ws['D1'] = 'Zone3'

    # Bold header styling
    header_font = Font(bold=True)
    for cell in ['B1', 'C1', 'D1']:
        ws[cell].font = header_font

    # --- Rate matrix data rows 2-5 ---
    # Row 2: Light
    ws['A2'] = 'Light'
    ws['B2'] = 5.99
    ws['C2'] = 8.99
    ws['D2'] = 12.99

    # Row 3: Medium
    ws['A3'] = 'Medium'
    ws['B3'] = 9.99
    ws['C3'] = 14.99
    ws['D3'] = 19.99

    # Row 4: Heavy
    ws['A4'] = 'Heavy'
    ws['B4'] = 15.99
    ws['C4'] = 22.99
    ws['D4'] = 32.99

    # Row 5: Oversized
    ws['A5'] = 'Oversized'
    ws['B5'] = 25.99
    ws['C5'] = 35.99
    ws['D5'] = 49.99

    # Bold row labels A2:A5
    for cell in ['A2', 'A3', 'A4', 'A5']:
        ws[cell].font = header_font

    # --- Row 6 blank separator ---

    # --- Lookup table header row 7 ---
    ws['A7'] = 'Weight Category'
    ws['B7'] = 'Destination Zone'
    ws['C7'] = 'Shipping Cost'

    lookup_header_font = Font(bold=True)
    for cell in ['A7', 'B7', 'C7']:
        ws[cell].font = lookup_header_font

    # --- Lookup data rows 8-17: weight categories and destination zones ---
    lookup_data = [
        ('Heavy',     'Zone2'),
        ('Light',     'Zone1'),
        ('Oversized', 'Zone3'),
        ('Medium',    'Zone2'),
        ('Light',     'Zone3'),
        ('Heavy',     'Zone1'),
        ('Medium',    'Zone1'),
        ('Oversized', 'Zone2'),
        ('Light',     'Zone2'),
        ('Heavy',     'Zone3'),
    ]

    for i, (weight, zone) in enumerate(lookup_data, start=8):
        ws.cell(row=i, column=1, value=weight)
        ws.cell(row=i, column=2, value=zone)
        # Column C (Shipping Cost) intentionally left empty

    # Column widths for readability
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
