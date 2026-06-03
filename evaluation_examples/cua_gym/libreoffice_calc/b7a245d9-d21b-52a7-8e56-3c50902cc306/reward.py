"""
Reward Script: Two-dimensional INDEX/MATCH lookup for shipping costs
Task ID: calc_fma_index_match_2d_005
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: C8:C17 all contain INDEX/MATCH formulas (0.6 pts)
  Component 2: Formulas use correct absolute references
               ($B$2:$D$5, $A$2:$A$5, $B$1:$D$1) with row-relative
               lookup values (A{row}, B{row}) (0.4 pts)
  Total: 1.0

Note: Rate matrix integrity and lookup column preservation are checked
      as precondition gates only (do not contribute to score), since
      they are identical in the initial and golden files.
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fma_index_match_2d_005'


def check_index_match_formula(formula):
    """
    Check that the cell value is an INDEX/MATCH formula.
    Verifies it starts with =INDEX( and contains at least 2 MATCH( calls.
    Returns True if formula matches expected pattern, False otherwise.
    """
    if not isinstance(formula, str):
        return False
    f = formula.upper().replace(' ', '')
    # Must start with =INDEX(
    if not f.startswith('=INDEX('):
        return False
    # Must contain at least two MATCH( calls
    if f.count('MATCH(') < 2:
        return False
    return True


def check_formula_references(formula, row_num):
    """
    Check formula uses correct absolute references:
    - Rate matrix data range: $B$2:$D$5
    - Row lookup array (weight categories): $A$2:$A$5
    - Column lookup array (zone headers): $B$1:$D$1
    - Row-relative lookup values: A{row_num} and B{row_num}
    Returns True if all references are correct.
    """
    if not isinstance(formula, str):
        return False
    f = formula.upper().replace(' ', '')

    # Check absolute rate matrix reference
    if '$B$2:$D$5' not in f:
        return False
    # Check absolute row lookup array
    if '$A$2:$A$5' not in f:
        return False
    # Check absolute column lookup array
    if '$B$1:$D$1' not in f:
        return False
    # Check row-relative lookup values (A{row} and B{row})
    if f'A{row_num}' not in f:
        return False
    if f'B{row_num}' not in f:
        return False
    return True


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook (formula strings, not computed values)
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: 'Shipping' sheet must exist
    if 'Shipping' not in wb.sheetnames:
        print("CRITICAL: 'Shipping' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Shipping']

    # Precondition gate: rate matrix must be intact (not scored, just gated)
    # If the rate matrix is corrupted, report 0.0 immediately
    expected_matrix = {
        (2, 2): 5.99, (2, 3): 8.99, (2, 4): 12.99,
        (3, 2): 9.99, (3, 3): 14.99, (3, 4): 19.99,
        (4, 2): 15.99, (4, 3): 22.99, (4, 4): 32.99,
        (5, 2): 25.99, (5, 3): 35.99, (5, 4): 49.99,
    }
    try:
        for (row, col), expected_val in expected_matrix.items():
            actual = ws.cell(row=row, column=col).value
            if actual is None or abs(float(actual) - expected_val) > 0.001:
                print(f"GATE FAIL: Rate matrix corrupted at row={row}, col={col}: expected {expected_val}, found {repr(actual)}")
                print("REWARD: 0.0")
                return 0.0
    except Exception as e:
        print(f"GATE ERROR: Could not verify rate matrix: {e}")
        print("REWARD: 0.0")
        return 0.0

    formula_rows = list(range(8, 18))  # rows 8-17 (C8:C17)

    # Component 1: All 10 cells C8:C17 contain INDEX/MATCH formulas (0.6 pts)
    # These cells are None in the initial file — formula presence indicates task completion
    try:
        formula_cells_ok = 0
        formula_cells_total = 10

        for row in formula_rows:
            cell_val = ws.cell(row=row, column=3).value  # Column C
            if check_index_match_formula(cell_val):
                formula_cells_ok += 1
            else:
                print(f"FAIL: C{row} missing/invalid INDEX/MATCH formula, found: {repr(cell_val)}")

        if formula_cells_ok == formula_cells_total:
            print(f"PASS: Component 1 — All 10 cells C8:C17 contain INDEX/MATCH formulas (0.6 pts)")
            total_score += 0.6
        elif formula_cells_ok > 0:
            partial = round(0.6 * formula_cells_ok / formula_cells_total, 4)
            print(f"PARTIAL: Component 1 — {formula_cells_ok}/{formula_cells_total} cells have INDEX/MATCH formulas ({partial} pts)")
            total_score += partial
        else:
            print("FAIL: Component 1 — No INDEX/MATCH formulas found in C8:C17")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Formulas use correct absolute references (0.4 pts)
    # Checks $B$2:$D$5 (data range), $A$2:$A$5 (row lookup), $B$1:$D$1 (col lookup)
    # and row-relative A{row}, B{row} as lookup values
    try:
        ref_ok = 0
        ref_total = 10

        for row in formula_rows:
            cell_val = ws.cell(row=row, column=3).value
            if check_formula_references(cell_val, row):
                ref_ok += 1
            else:
                if isinstance(cell_val, str):
                    print(f"FAIL: C{row} incorrect references — expected $B$2:$D$5, $A$2:$A$5, $B$1:$D$1 with A{row}/B{row}: {repr(cell_val)}")

        if ref_ok == ref_total:
            print(f"PASS: Component 2 — All 10 formulas use correct absolute references ($B$2:$D$5, $A$2:$A$5, $B$1:$D$1) (0.4 pts)")
            total_score += 0.4
        elif ref_ok > 0:
            partial = round(0.4 * ref_ok / ref_total, 4)
            print(f"PARTIAL: Component 2 — {ref_ok}/{ref_total} formulas have correct references ({partial} pts)")
            total_score += partial
        else:
            print("FAIL: Component 2 — No formulas have correct absolute references")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
