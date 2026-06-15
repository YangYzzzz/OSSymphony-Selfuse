"""
Reward Script: NETWORKDAYS formula in attendance spreadsheet
Task ID: calc_gg5_022
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Majority of F2:F61 contain NETWORKDAYS formulas
  Component 2 (0.3): Formulas reference Holidays.$A$2:$A$15 with absolute refs
  Component 3 (0.3): Formulas use correct row-relative C/D cell references
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_022'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Monthly' sheet must exist
    if 'Monthly' not in wb.sheetnames:
        print("FAIL: 'Monthly' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Monthly']

    # We expect formulas in F2:F61 (60 employee rows)
    total_rows = 60  # rows 2 through 61
    networkdays_count = 0
    holidays_ref_correct = 0
    row_ref_correct = 0

    for r in range(2, 62):
        cell_val = ws.cell(row=r, column=6).value  # Column F
        if cell_val is None:
            continue

        val_str = str(cell_val).strip().upper().replace(" ", "")

        # Check if it contains NETWORKDAYS function
        if "NETWORKDAYS(" in val_str:
            networkdays_count += 1

            # Check if holidays reference uses absolute refs to Holidays sheet
            # Expected pattern: Holidays.$A$2:$A$15 (case-insensitive)
            if "HOLIDAYS.$A$2:$A$15" in val_str:
                holidays_ref_correct += 1

            # Check if row-relative C and D references are correct for this row
            # Expected: C<r> and D<r> (e.g., C2 and D2 for row 2)
            expected_c = f"C{r}"
            expected_d = f"D{r}"
            if expected_c in val_str and expected_d in val_str:
                row_ref_correct += 1

    print(f"INFO: Found {networkdays_count}/{total_rows} NETWORKDAYS formulas")
    print(f"INFO: {holidays_ref_correct}/{total_rows} with correct Holidays ref")
    print(f"INFO: {row_ref_correct}/{total_rows} with correct row-relative refs")

    # Component 1: Majority of F2:F61 contain NETWORKDAYS formulas (0.4 points)
    # Must have at least 80% (48/60) to get full credit, partial for 50%+
    try:
        ratio = networkdays_count / total_rows
        if ratio >= 0.8:
            print(f"PASS: Component 1 — {networkdays_count}/{total_rows} NETWORKDAYS formulas ({ratio:.0%}) (0.4 pts)")
            total_score += 0.4
        elif ratio >= 0.5:
            partial = 0.4 * (ratio - 0.5) / 0.3  # linear scale from 50% to 80%
            print(f"PARTIAL: Component 1 — {networkdays_count}/{total_rows} NETWORKDAYS formulas ({ratio:.0%}) ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — only {networkdays_count}/{total_rows} NETWORKDAYS formulas ({ratio:.0%})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Formulas reference Holidays.$A$2:$A$15 correctly (0.3 points)
    # Only score if we found NETWORKDAYS formulas
    try:
        if networkdays_count > 0:
            ref_ratio = holidays_ref_correct / networkdays_count
            if ref_ratio >= 0.8:
                print(f"PASS: Component 2 — {holidays_ref_correct}/{networkdays_count} correct Holidays refs ({ref_ratio:.0%}) (0.3 pts)")
                total_score += 0.3
            elif ref_ratio >= 0.5:
                partial = 0.3 * (ref_ratio - 0.5) / 0.3
                print(f"PARTIAL: Component 2 — {holidays_ref_correct}/{networkdays_count} correct Holidays refs ({partial:.2f} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 2 — only {holidays_ref_correct}/{networkdays_count} correct Holidays refs")
        else:
            print(f"FAIL: Component 2 — no NETWORKDAYS formulas to check refs against")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Formulas use correct row-relative C/D references (0.3 points)
    try:
        if networkdays_count > 0:
            row_ratio = row_ref_correct / networkdays_count
            if row_ratio >= 0.8:
                print(f"PASS: Component 3 — {row_ref_correct}/{networkdays_count} correct row refs ({row_ratio:.0%}) (0.3 pts)")
                total_score += 0.3
            elif row_ratio >= 0.5:
                partial = 0.3 * (row_ratio - 0.5) / 0.3
                print(f"PARTIAL: Component 3 — {row_ref_correct}/{networkdays_count} correct row refs ({partial:.2f} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 3 — only {row_ref_correct}/{networkdays_count} correct row refs")
        else:
            print(f"FAIL: Component 3 — no NETWORKDAYS formulas to check row refs against")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
