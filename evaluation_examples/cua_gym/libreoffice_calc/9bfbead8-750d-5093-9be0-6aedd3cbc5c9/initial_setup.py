"""
Initial Setup: Add NETWORKDAYS formulas to attendance spreadsheet
Task ID: calc_gg5_022
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_022'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Monthly ---
    ws = wb.active
    ws.title = 'Monthly'

    # Header row styling
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    headers = ['Employee ID', 'Name', 'Period Start', 'Period End', 'Sick Days', 'Working Days']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14

    # Employee data - 60 rows of realistic content
    first_names = [
        'Sarah', 'Marcus', 'Priya', 'James', 'Mei', 'Carlos', 'Fatima', 'David',
        'Yuki', 'Oliver', 'Amara', 'Liam', 'Sofia', 'Chen', 'Elena', 'Michael',
        'Aisha', 'Robert', 'Hana', 'Thomas', 'Zara', 'William', 'Ling', 'Daniel',
        'Nadia', 'Patrick', 'Rosa', 'Kevin', 'Ingrid', 'Samuel', 'Keiko', 'Frank',
        'Valentina', 'George', 'Layla', 'Peter', 'Ananya', 'Steven', 'Marta', 'Andrew',
        'Chloe', 'Ryan', 'Devi', 'Nathan', 'Isabel', 'Brian', 'Sunita', 'Mark',
        'Freya', 'Jason', 'Neha', 'Timothy', 'Olivia', 'Derek', 'Maya', 'Vincent',
        'Tara', 'Lucas', 'Rina', 'Edward'
    ]
    last_names = [
        'Chen', 'Johnson', 'Patel', 'O\'Brien', 'Wang', 'Rodriguez', 'Al-Rashid', 'Kim',
        'Tanaka', 'Smith', 'Okafor', 'Murphy', 'Hernandez', 'Wei', 'Petrov', 'Thompson',
        'Ibrahim', 'Williams', 'Nakamura', 'Anderson', 'Hassan', 'Brown', 'Zhang', 'Taylor',
        'Kowalski', 'O\'Connor', 'Morales', 'Lee', 'Bergstrom', 'Davis', 'Sato', 'Weber',
        'Rossi', 'Harris', 'Khoury', 'Muller', 'Sharma', 'Clark', 'Gonzalez', 'Wilson',
        'Dubois', 'Mitchell', 'Gupta', 'Hall', 'Santos', 'Cooper', 'Reddy', 'Evans',
        'Johansson', 'Martin', 'Kapoor', 'Baker', 'Bennett', 'Price', 'Iyer', 'Tran',
        'Singh', 'Adams', 'Fernandez', 'Wright'
    ]

    import random
    random.seed(42)  # Reproducible data

    base_start = date(2025, 1, 6)  # First Monday of January 2025

    for i in range(60):
        row = i + 2
        emp_id = f'EMP-{1001 + i:04d}'
        name = f'{first_names[i]} {last_names[i]}'

        # Each employee has a period within Jan-Mar 2025
        month_offset = i % 3
        period_start = date(2025, 1 + month_offset, 1 + (i % 20))
        period_end = period_start + timedelta(days=20 + (i % 10))
        # Clamp end to reasonable date
        if period_end > date(2025, 4, 30):
            period_end = date(2025, 4, 30)

        sick_days = random.randint(0, 5)

        ws.cell(row=row, column=1, value=emp_id)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=period_start)
        ws.cell(row=row, column=3).number_format = 'yyyy-mm-dd'
        ws.cell(row=row, column=4, value=period_end)
        ws.cell(row=row, column=4).number_format = 'yyyy-mm-dd'
        ws.cell(row=row, column=5, value=sick_days)
        # Column F (Working Days) intentionally left EMPTY

    # Freeze header row
    ws.freeze_panes = 'A2'

    # --- Sheet 2: Holidays ---
    ws_holidays = wb.create_sheet('Holidays')
    ws_holidays.cell(row=1, column=1, value='Holiday Date')
    ws_holidays['A1'].font = Font(name='Calibri', size=11, bold=True)
    ws_holidays.column_dimensions['A'].width = 16

    # 14 US public holidays for 2025
    holidays = [
        date(2025, 1, 1),    # New Year's Day
        date(2025, 1, 20),   # Martin Luther King Jr. Day
        date(2025, 2, 17),   # Presidents' Day
        date(2025, 3, 31),   # Cesar Chavez Day (observed)
        date(2025, 5, 26),   # Memorial Day
        date(2025, 6, 19),   # Juneteenth
        date(2025, 7, 4),    # Independence Day
        date(2025, 9, 1),    # Labor Day
        date(2025, 10, 13),  # Columbus Day
        date(2025, 11, 11),  # Veterans Day
        date(2025, 11, 27),  # Thanksgiving
        date(2025, 11, 28),  # Day after Thanksgiving
        date(2025, 12, 25),  # Christmas Day
        date(2025, 12, 31),  # New Year's Eve
    ]

    for idx, hol in enumerate(holidays):
        cell = ws_holidays.cell(row=idx + 2, column=1, value=hol)
        cell.number_format = 'yyyy-mm-dd'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
