"""
Initial Setup: Product Launch Project Checklist
Task ID: calc_grs_058
Domain: libreoffice_calc

Creates a multi-department product launch checklist with raw data only.
No conditional formatting, no dropdowns, no summary formulas, no critical path view.
"""

import os
import shlex
import subprocess
import time
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_058'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # Common styles for headers
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    headers = ["Task Description", "Owner", "Due Date", "Status", "Priority", "Dependencies"]

    # Base date for due dates (launch date = 6 weeks from a reference)
    launch_date = date(2026, 5, 15)

    def write_headers(ws):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 30
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30

    def write_data(ws, data):
        for r, row_data in enumerate(data, 2):
            for c, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if c == 3 and isinstance(val, date):
                    cell.number_format = 'yyyy-mm-dd'

    # ============================================================
    # Sheet 1: Marketing (10 tasks)
    # ============================================================
    ws_mkt = wb.active
    ws_mkt.title = "Marketing"
    write_headers(ws_mkt)

    marketing_data = [
        ["Develop launch messaging and positioning document", "Sarah Chen", date(2026, 4, 4), "In Progress", "High", "None"],
        ["Create product demo video (2-min overview)", "James Rivera", date(2026, 4, 18), "Not Started", "High", "Messaging doc complete"],
        ["Design landing page wireframes and copy", "Priya Sharma", date(2026, 4, 11), "In Progress", "High", "Messaging doc complete"],
        ["Write blog post announcing the feature", "Sarah Chen", date(2026, 4, 25), "Not Started", "Medium", "Landing page approved"],
        ["Prepare email campaign (3-touch sequence)", "Marcus Johnson", date(2026, 4, 22), "Not Started", "Medium", "Messaging doc complete"],
        ["Create social media assets (LinkedIn, Twitter, Instagram)", "Aisha Williams", date(2026, 4, 25), "Not Started", "Medium", "Landing page live"],
        ["Brief analyst relations team", "David Park", date(2026, 5, 2), "Not Started", "Low", "Messaging doc complete"],
        ["Set up launch day webinar logistics", "James Rivera", date(2026, 5, 5), "Not Started", "High", "Demo video complete"],
        ["Coordinate press release with PR agency", "Sarah Chen", date(2026, 5, 8), "Not Started", "Medium", "Messaging doc approved"],
        ["Launch day social media blitz execution plan", "Aisha Williams", date(2026, 5, 12), "Not Started", "High", "Social assets ready"],
    ]
    write_data(ws_mkt, marketing_data)

    # ============================================================
    # Sheet 2: Product (8 tasks)
    # ============================================================
    ws_prod = wb.create_sheet("Product")
    write_headers(ws_prod)

    product_data = [
        ["Complete feature development and code freeze", "Alex Thompson", date(2026, 4, 7), "In Progress", "High", "None"],
        ["Internal QA testing round 1", "Wei Zhang", date(2026, 4, 14), "Not Started", "High", "Code freeze"],
        ["Fix critical bugs from QA round 1", "Alex Thompson", date(2026, 4, 18), "Not Started", "High", "QA round 1 complete"],
        ["Beta testing with 5 key customer accounts", "Rachel Kim", date(2026, 4, 25), "Not Started", "High", "QA round 1 fixes"],
        ["Incorporate beta feedback and final polish", "Alex Thompson", date(2026, 5, 2), "Not Started", "Medium", "Beta complete"],
        ["Write internal knowledge base articles", "Wei Zhang", date(2026, 4, 28), "Not Started", "Medium", "Feature stable"],
        ["Performance and load testing", "Tomasz Kowalski", date(2026, 5, 5), "Not Started", "High", "Final polish done"],
        ["Go/No-Go release review with engineering leads", "Rachel Kim", date(2026, 5, 9), "Not Started", "High", "Load testing pass"],
    ]
    write_data(ws_prod, product_data)

    # ============================================================
    # Sheet 3: Sales Enablement (8 tasks)
    # ============================================================
    ws_sales = wb.create_sheet("Sales Enablement")
    write_headers(ws_sales)

    sales_data = [
        ["Create sales battlecard with competitive positioning", "Jordan Mitchell", date(2026, 4, 11), "Not Started", "High", "Messaging doc complete"],
        ["Develop ROI calculator for prospects", "Natalie Flores", date(2026, 4, 18), "Not Started", "Medium", "Feature specs finalized"],
        ["Build feature comparison matrix vs competitors", "Jordan Mitchell", date(2026, 4, 14), "Not Started", "High", "None"],
        ["Record sales training webinar (45 min)", "Chris Dawson", date(2026, 4, 25), "Not Started", "High", "Battlecard ready"],
        ["Update CRM with new product fields and tags", "Natalie Flores", date(2026, 4, 21), "Not Started", "Medium", "None"],
        ["Create objection handling guide", "Jordan Mitchell", date(2026, 4, 28), "Not Started", "Medium", "Battlecard ready"],
        ["Prepare customer case study from beta participants", "Chris Dawson", date(2026, 5, 5), "Not Started", "Low", "Beta complete"],
        ["Conduct live sales team training session", "Jordan Mitchell", date(2026, 5, 9), "Not Started", "High", "Training webinar recorded"],
    ]
    write_data(ws_sales, sales_data)

    # ============================================================
    # Sheet 4: Operations (6 tasks)
    # ============================================================
    ws_ops = wb.create_sheet("Operations")
    write_headers(ws_ops)

    ops_data = [
        ["Provision production infrastructure and scaling", "Daniel Okafor", date(2026, 4, 14), "Not Started", "High", "Code freeze"],
        ["Configure monitoring dashboards and alerts", "Elena Vasquez", date(2026, 4, 21), "Not Started", "High", "Infrastructure ready"],
        ["Update runbook and incident response procedures", "Daniel Okafor", date(2026, 4, 28), "Not Started", "Medium", "Monitoring configured"],
        ["Conduct load test on production environment", "Elena Vasquez", date(2026, 5, 2), "Not Started", "High", "Infrastructure ready"],
        ["Set up feature flags for staged rollout", "Daniel Okafor", date(2026, 4, 25), "Not Started", "Medium", "Code freeze"],
        ["Coordinate with support team on escalation paths", "Elena Vasquez", date(2026, 5, 5), "Not Started", "Low", "Runbook updated"],
    ]
    write_data(ws_ops, ops_data)

    # ============================================================
    # Sheet 5: Legal Compliance (5 tasks)
    # ============================================================
    ws_legal = wb.create_sheet("Legal Compliance")
    write_headers(ws_legal)

    legal_data = [
        ["Review updated Terms of Service for new feature", "Michael Torres", date(2026, 4, 7), "In Progress", "High", "None"],
        ["Complete data privacy impact assessment (DPIA)", "Sana Patel", date(2026, 4, 14), "Not Started", "High", "Feature specs finalized"],
        ["Update privacy policy and cookie consent flows", "Michael Torres", date(2026, 4, 21), "Not Started", "High", "DPIA complete"],
        ["Obtain SOC 2 compliance sign-off for new data flows", "Sana Patel", date(2026, 4, 28), "Not Started", "Medium", "DPIA complete"],
        ["Final legal review and launch approval sign-off", "Michael Torres", date(2026, 5, 9), "Not Started", "High", "All compliance items done"],
    ]
    write_data(ws_legal, legal_data)

    # ============================================================
    # Sheet 6: Summary (placeholder - no formulas yet)
    # ============================================================
    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "Department"
    ws_summary["B1"] = "Total Tasks"
    ws_summary["C1"] = "Completed"
    ws_summary["D1"] = "Completion %"

    for col in range(1, 5):
        cell = ws_summary.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    departments = ["Marketing", "Product", "Sales Enablement", "Operations", "Legal Compliance"]
    task_counts = [10, 8, 8, 6, 5]
    for r, (dept, count) in enumerate(zip(departments, task_counts), 2):
        ws_summary.cell(row=r, column=1, value=dept).border = thin_border
        ws_summary.cell(row=r, column=2, value=count).border = thin_border
        # Columns C and D intentionally left empty - agent must add formulas
        ws_summary.cell(row=r, column=3).border = thin_border
        ws_summary.cell(row=r, column=4).border = thin_border

    # Row 7: Launch Readiness Score label (no formula)
    ws_summary.cell(row=7, column=1, value="Launch Readiness Score").font = Font(bold=True, size=12)

    ws_summary.column_dimensions["A"].width = 22
    ws_summary.column_dimensions["B"].width = 14
    ws_summary.column_dimensions["C"].width = 14
    ws_summary.column_dimensions["D"].width = 16

    # ============================================================
    # Sheet 7: Critical Path (headers only, no data)
    # ============================================================
    ws_cp = wb.create_sheet("Critical Path")
    cp_headers = ["Task Description", "Department", "Owner", "Due Date", "Status", "Dependencies"]
    for col, h in enumerate(cp_headers, 1):
        cell = ws_cp.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color="FFC00000", end_color="FFC00000", fill_type="solid")
        cell.alignment = header_align
        cell.border = thin_border

    ws_cp.column_dimensions["A"].width = 45
    ws_cp.column_dimensions["B"].width = 20
    ws_cp.column_dimensions["C"].width = 20
    ws_cp.column_dimensions["D"].width = 14
    ws_cp.column_dimensions["E"].width = 14
    ws_cp.column_dimensions["F"].width = 30
    ws_cp.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
