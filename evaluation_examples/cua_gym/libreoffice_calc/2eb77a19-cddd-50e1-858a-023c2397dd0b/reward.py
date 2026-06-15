"""
Reward Script: Product Launch Project Checklist
Task ID: calc_grs_058
Domain: libreoffice_calc
Scoring:
  Component 1: Data validations on department sheets (0.20)
  Component 2: Conditional formatting on department sheets (0.20)
  Component 3: Summary sheet formulas (COUNTIF + completion %) (0.25)
  Component 4: Launch Readiness Score in Summary (0.10)
  Component 5: Critical Path sheet populated with High Priority items sorted by date (0.25)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_058'

DEPT_SHEETS = ['Marketing', 'Product', 'Sales Enablement', 'Operations', 'Legal Compliance']
DEPT_TASK_COUNTS = {'Marketing': 10, 'Product': 8, 'Sales Enablement': 8, 'Operations': 6, 'Legal Compliance': 5}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Data validations on department sheets (0.20 points)
    # Task requires Status dropdown (Not Started, In Progress, Review, Done, Blocked)
    # and Priority dropdown (High, Medium, Low) on each department sheet.
    # Initial file has NO data validations; golden adds them.
    try:
        dv_score = 0.0
        sheets_with_both_dv = 0
        for dept in DEPT_SHEETS:
            if dept not in wb.sheetnames:
                continue
            ws = wb[dept]
            dvs = ws.data_validations.dataValidation if ws.data_validations else []
            has_status_dv = False
            has_priority_dv = False
            for dv in dvs:
                if dv.type == 'list' and dv.formula1:
                    formula_lower = dv.formula1.lower().replace('"', '').replace("'", '')
                    if 'done' in formula_lower and 'not started' in formula_lower:
                        has_status_dv = True
                    if 'high' in formula_lower and 'medium' in formula_lower and 'low' in formula_lower:
                        has_priority_dv = True
            if has_status_dv and has_priority_dv:
                sheets_with_both_dv += 1

        if sheets_with_both_dv >= 5:
            dv_score = 0.20
            print(f"PASS: Component 1 — All 5 department sheets have Status+Priority dropdowns (0.20 pts)")
        elif sheets_with_both_dv >= 3:
            dv_score = 0.12
            print(f"PARTIAL: Component 1 — {sheets_with_both_dv}/5 dept sheets have both dropdowns (0.12 pts)")
        elif sheets_with_both_dv >= 1:
            dv_score = 0.06
            print(f"PARTIAL: Component 1 — {sheets_with_both_dv}/5 dept sheets have both dropdowns (0.06 pts)")
        else:
            print(f"FAIL: Component 1 — No department sheets have Status+Priority dropdowns")

        total_score += dv_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Conditional formatting on department sheets (0.20 points)
    # Task requires status color CF and overdue highlighting (due < today and not Done).
    # Initial file has NO conditional formatting; golden adds it.
    try:
        cf_score = 0.0
        sheets_with_cf = 0
        for dept in DEPT_SHEETS:
            if dept not in wb.sheetnames:
                continue
            ws = wb[dept]
            cf_rules = list(ws.conditional_formatting)
            # Golden has 2 CF rule ranges per dept sheet: status colors + overdue
            if len(cf_rules) >= 2:
                sheets_with_cf += 1
            elif len(cf_rules) >= 1:
                sheets_with_cf += 0.5

        if sheets_with_cf >= 4.5:
            cf_score = 0.20
            print(f"PASS: Component 2 — All 5 dept sheets have conditional formatting (0.20 pts)")
        elif sheets_with_cf >= 2.5:
            cf_score = 0.12
            print(f"PARTIAL: Component 2 — {sheets_with_cf}/5 dept sheets have CF (0.12 pts)")
        elif sheets_with_cf >= 1:
            cf_score = 0.06
            print(f"PARTIAL: Component 2 — {sheets_with_cf}/5 dept sheets have CF (0.06 pts)")
        else:
            print(f"FAIL: Component 2 — No department sheets have conditional formatting")

        total_score += cf_score
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Summary sheet with COUNTIF formulas and completion percentages (0.25 points)
    # Initial Summary has department names and Total Tasks but C and D columns are empty.
    # Golden adds COUNTIF formulas in C2:C6, IF formulas in D2:D6, and a Total row (row 7).
    try:
        summary_score = 0.0
        if 'Summary' not in wb.sheetnames:
            print(f"FAIL: Component 3 — 'Summary' sheet not found")
        else:
            ws = wb['Summary']
            # Check that C2:C6 have COUNTIF formulas
            countif_count = 0
            for row_idx in range(2, 7):
                val = ws.cell(row=row_idx, column=3).value
                if val and isinstance(val, str) and 'COUNTIF' in val.upper():
                    countif_count += 1

            # Check that D2:D6 have percentage/completion formulas
            pct_count = 0
            for row_idx in range(2, 7):
                val = ws.cell(row=row_idx, column=4).value
                if val and isinstance(val, str) and ('/' in val or 'IF' in val.upper()):
                    pct_count += 1

            # Check Total row exists (row 7 with SUM formulas)
            total_row_ok = False
            b7 = ws.cell(row=7, column=2).value
            c7 = ws.cell(row=7, column=3).value
            if b7 and isinstance(b7, str) and 'SUM' in b7.upper():
                total_row_ok = True
            elif c7 and isinstance(c7, str) and 'SUM' in c7.upper():
                total_row_ok = True

            if countif_count >= 5 and pct_count >= 5 and total_row_ok:
                summary_score = 0.25
                print(f"PASS: Component 3 — Summary has COUNTIF formulas ({countif_count}), "
                      f"completion % ({pct_count}), Total row present (0.25 pts)")
            elif countif_count >= 3 and pct_count >= 3:
                summary_score = 0.15
                print(f"PARTIAL: Component 3 — COUNTIF={countif_count}/5, PCT={pct_count}/5, "
                      f"Total={total_row_ok} (0.15 pts)")
            elif countif_count >= 1 or pct_count >= 1:
                summary_score = 0.08
                print(f"PARTIAL: Component 3 — COUNTIF={countif_count}/5, PCT={pct_count}/5, "
                      f"Total={total_row_ok} (0.08 pts)")
            else:
                print(f"FAIL: Component 3 — No COUNTIF or completion formulas in Summary")

            total_score += summary_score
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Launch Readiness Score (0.10 points)
    # Initial file has no readiness score. Golden has a formula in B9 referencing overall completion.
    try:
        readiness_score = 0.0
        if 'Summary' in wb.sheetnames:
            ws = wb['Summary']
            # Check for "Launch Readiness Score" label and formula
            found_label = False
            found_formula = False
            # Search rows 7-15 for the label
            for row_idx in range(7, min(ws.max_row + 1, 16)):
                a_val = ws.cell(row=row_idx, column=1).value
                if a_val and isinstance(a_val, str) and 'readiness' in a_val.lower():
                    found_label = True
                    # Check adjacent cell for formula
                    b_val = ws.cell(row=row_idx, column=2).value
                    d_val = ws.cell(row=row_idx, column=4).value
                    for check_val in [b_val, d_val]:
                        if check_val and isinstance(check_val, str) and '=' in check_val:
                            found_formula = True
                            break

            if found_label and found_formula:
                readiness_score = 0.10
                print(f"PASS: Component 4 — Launch Readiness Score label + formula found (0.10 pts)")
            else:
                print(f"FAIL: Component 4 — No Launch Readiness Score found in Summary")
        else:
            print(f"FAIL: Component 4 — Summary sheet not found")

        total_score += readiness_score
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Critical Path view with High Priority items sorted by due date (0.25 points)
    # Initial Critical Path sheet has only headers (row 1). Golden populates it with
    # High Priority tasks from all departments, sorted by due date.
    try:
        cp_score = 0.0
        if 'Critical Path' not in wb.sheetnames:
            print(f"FAIL: Component 5 — 'Critical Path' sheet not found")
        else:
            ws = wb['Critical Path']
            # Count data rows (rows beyond header)
            data_rows = 0
            for row_idx in range(2, ws.max_row + 1):
                a_val = ws.cell(row=row_idx, column=1).value
                if a_val and str(a_val).strip():
                    data_rows += 1

            # Check sorting by due date (column D should be non-decreasing)
            dates_sorted = True
            prev_date = None
            for row_idx in range(2, ws.max_row + 1):
                d_val = ws.cell(row=row_idx, column=4).value
                if d_val is None:
                    continue
                try:
                    import datetime
                    if isinstance(d_val, datetime.datetime):
                        curr_date = d_val
                    else:
                        continue
                    if prev_date and curr_date < prev_date:
                        dates_sorted = False
                    prev_date = curr_date
                except:
                    pass

            # Check that items include a Department column (column B)
            has_dept_col = False
            header_b = ws.cell(row=1, column=2).value
            if header_b and 'department' in str(header_b).lower():
                has_dept_col = True

            # Check CF for overdue highlighting exists
            has_cf = len(list(ws.conditional_formatting)) > 0

            if data_rows >= 15 and dates_sorted and has_dept_col:
                cp_score = 0.25
                print(f"PASS: Component 5 — Critical Path has {data_rows} rows, "
                      f"sorted={dates_sorted}, dept_col={has_dept_col}, CF={has_cf} (0.25 pts)")
            elif data_rows >= 10 and dates_sorted:
                cp_score = 0.18
                print(f"PARTIAL: Component 5 — {data_rows} rows, sorted, dept_col={has_dept_col} (0.18 pts)")
            elif data_rows >= 5:
                cp_score = 0.10
                print(f"PARTIAL: Component 5 — {data_rows} rows populated (0.10 pts)")
            else:
                print(f"FAIL: Component 5 — Critical Path has only {data_rows} data rows (need >= 5)")

            total_score += cp_score
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook for LibreOffice
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Main execution
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'

# Try to persist unsaved state
persist_app_state()

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
