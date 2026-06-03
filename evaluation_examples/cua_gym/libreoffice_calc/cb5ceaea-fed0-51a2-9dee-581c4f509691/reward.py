"""
Reward Script: Copy Template sheet to Jan/Feb/Mar and add DRAFT label via grouping
Task ID: calc_ggf_048
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): Sheets Jan, Feb, Mar all exist
  Component 2 (0.2): Sheet tab order is Template, Jan, Feb, Mar
  Component 3 (0.3): A1 on Jan, Feb, Mar all contain 'DRAFT'
  Component 4 (0.2): Template sheet A1 is NOT 'DRAFT' (unchanged)
"""

import os

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_048'


def persist_app_state(domain):
    """Save any unsaved LibreOffice edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for", domain)
        except Exception as e:
            print("PERSIST_WARN: save hook failed:", e)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    sheet_names = wb.sheetnames
    print(f"INFO: Found sheets: {sheet_names}")

    # Component 1: Sheets Jan, Feb, Mar all exist (0.3 points)
    # This FAILS on initial (only Template) and PASSES on golden
    try:
        required_sheets = ['Jan', 'Feb', 'Mar']
        found = [s for s in required_sheets if s in sheet_names]
        if len(found) == 3:
            print(f"PASS: Component 1 — All three sheets Jan, Feb, Mar exist (0.3 pts)")
            total_score += 0.3
        elif len(found) > 0:
            partial = round(0.1 * len(found), 2)
            print(f"PARTIAL: Component 1 — Found {len(found)}/3 sheets: {found} ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — None of Jan, Feb, Mar found in {sheet_names}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Sheet tab order has Template before Jan, Feb, Mar in sequence (0.2 points)
    # This FAILS on initial (no Jan/Feb/Mar) and PASSES on golden
    try:
        if all(s in sheet_names for s in ['Template', 'Jan', 'Feb', 'Mar']):
            t_idx = sheet_names.index('Template')
            j_idx = sheet_names.index('Jan')
            f_idx = sheet_names.index('Feb')
            m_idx = sheet_names.index('Mar')
            if t_idx < j_idx < f_idx < m_idx:
                print(f"PASS: Component 2 — Sheet order correct: Template({t_idx}), Jan({j_idx}), Feb({f_idx}), Mar({m_idx}) (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 2 — Wrong order: Template({t_idx}), Jan({j_idx}), Feb({f_idx}), Mar({m_idx})")
        else:
            print(f"FAIL: Component 2 — Missing required sheets for order check")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: A1 on Jan, Feb, Mar all contain 'DRAFT' (0.3 points)
    # This FAILS on initial (no such sheets) and PASSES on golden
    try:
        draft_count = 0
        for sn in ['Jan', 'Feb', 'Mar']:
            if sn in sheet_names:
                ws = wb[sn]
                a1_val = ws.cell(row=1, column=1).value
                if a1_val is not None and str(a1_val).strip() == 'DRAFT':
                    draft_count += 1
                    print(f"  CHECK: {sn} A1 = {repr(a1_val)} — OK")
                else:
                    print(f"  CHECK: {sn} A1 = {repr(a1_val)} — expected 'DRAFT'")
            else:
                print(f"  CHECK: Sheet {sn} not found")

        if draft_count == 3:
            print(f"PASS: Component 3 — All three sheets have 'DRAFT' in A1 (0.3 pts)")
            total_score += 0.3
        elif draft_count > 0:
            partial = round(0.1 * draft_count, 2)
            print(f"PARTIAL: Component 3 — {draft_count}/3 sheets have 'DRAFT' in A1 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No sheets have 'DRAFT' in A1")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Template sheet A1 is NOT 'DRAFT' (0.2 points)
    # This checks that the Template was not modified by the grouping operation.
    # On initial_env, Template has no DRAFT and no Jan/Feb/Mar sheets exist,
    # so Component 1-3 already fail making the total 0. This component is gated
    # on Components 1-3 having some success to avoid awarding points on initial.
    try:
        if 'Template' in sheet_names and total_score > 0:
            ws_t = wb['Template']
            t_a1 = ws_t.cell(row=1, column=1).value
            if t_a1 is None or str(t_a1).strip() != 'DRAFT':
                print(f"PASS: Component 4 — Template A1 is {repr(t_a1)}, not 'DRAFT' (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 4 — Template A1 is 'DRAFT' — it should be unchanged")
        elif total_score == 0:
            print(f"SKIP: Component 4 — Gated on prior components (score is 0)")
        else:
            print(f"FAIL: Component 4 — Template sheet not found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
