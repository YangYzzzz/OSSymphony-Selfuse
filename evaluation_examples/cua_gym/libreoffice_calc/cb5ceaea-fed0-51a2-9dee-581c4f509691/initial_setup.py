"""
Initial Setup: Copy Template sheet and group-edit task
Task ID: calc_ggf_048
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_048'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Template'

    # --- Styles ---
    header_font = Font(name="Arial", size=14, bold=True, color="1F4E79")
    header_fill = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")
    col_header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    col_header_fill = PatternFill(start_color="FF2E75B6", end_color="FF2E75B6", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    category_font = Font(name="Arial", size=11)
    total_font = Font(name="Arial", size=11, bold=True)

    # --- Row 1: Report Title (merged) ---
    ws.merge_cells("A1:F1")
    ws["A1"] = "Monthly Report"
    ws["A1"].font = header_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = header_fill
    ws.row_dimensions[1].height = 30

    # --- Row 2: blank spacer ---
    ws.row_dimensions[2].height = 8

    # --- Row 3: Column Headers ---
    col_headers = ["Category", "Week 1", "Week 2", "Week 3", "Week 4", "Total"]
    for c, h in enumerate(col_headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # --- Rows 4-13: Category rows (empty data, just labels) ---
    categories = [
        "Revenue",
        "Cost of Goods Sold",
        "Gross Profit",
        "Payroll",
        "Marketing",
        "Office Supplies",
        "Utilities",
        "Insurance",
        "Travel & Entertainment",
        "Miscellaneous",
    ]
    for r, cat in enumerate(categories, 4):
        cell = ws.cell(row=r, column=1, value=cat)
        cell.font = category_font
        cell.border = thin_border
        # Add borders on data columns (empty cells)
        for c in range(2, 7):
            data_cell = ws.cell(row=r, column=c)
            data_cell.border = thin_border
            data_cell.alignment = center_align
            data_cell.number_format = '#,##0.00'

    # --- Row 14: Totals row ---
    total_row = 14
    cell = ws.cell(row=total_row, column=1, value="Total")
    cell.font = total_font
    cell.border = thin_border
    for c in range(2, 7):
        data_cell = ws.cell(row=total_row, column=c)
        data_cell.border = thin_border
        data_cell.font = total_font
        data_cell.alignment = center_align
        data_cell.number_format = '#,##0.00'

    # --- Row 16: Notes section ---
    ws.cell(row=16, column=1, value="Notes:").font = Font(name="Arial", size=11, bold=True, italic=True)
    ws.merge_cells("A17:F19")
    ws["A17"] = ""

    # --- Column widths ---
    ws.column_dimensions["A"].width = 22
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 14

    # --- Freeze panes at row 3 ---
    ws.freeze_panes = "A4"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
