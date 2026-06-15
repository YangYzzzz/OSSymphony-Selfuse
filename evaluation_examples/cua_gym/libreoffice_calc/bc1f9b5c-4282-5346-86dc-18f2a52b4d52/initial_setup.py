"""
Initial Setup: Financial report spreadsheet with decimal data
Task ID: osworld_calc_decimal_separator_005
Domain: libreoffice_calc

Creates a financial report spreadsheet with realistic financial data.
Cell A1 is initially used as a label for the report (not the task-completion value).
The agent must toggle decimal separator settings and document the final state in A1.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_decimal_separator_005'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Financial Report ---
    ws = wb.active
    ws.title = "Financial Report"

    # Header styles
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2E4057", end_color="FF2E4057", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1 — report title (NOT the task completion text)
    ws.merge_cells("A1:G1")
    ws["A1"] = "Meridian Capital — Regional Financial Report 2025"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill(start_color="FF1A2B3C", end_color="FF1A2B3C", fill_type="solid")
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws.row_dimensions[1].height = 32

    # Row 2 — blank spacer
    ws.row_dimensions[2].height = 8

    # Row 3 — column headers
    headers = ["Region", "Q1 Revenue", "Q2 Revenue", "Q3 Revenue", "Q4 Revenue", "Annual Total", "YoY Change (%)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    ws.row_dimensions[3].height = 22

    # Realistic financial data (rows 4–16)
    data = [
        ["North America",    1245830.75, 1389240.50, 1502310.25, 1678450.00, None, None],
        ["Western Europe",    987320.40,  1043180.60, 1124760.80,  1256890.30, None, None],
        ["Asia Pacific",     1534670.20, 1712340.90, 1893450.15,  2105630.75, None, None],
        ["Latin America",     423180.55,  467920.35,  512340.70,   589760.45, None, None],
        ["Middle East",       312450.80,  298760.25,  334510.60,   367240.90, None, None],
        ["Eastern Europe",    234780.15,  267310.40,  289450.75,   321680.20, None, None],
        ["Sub-Saharan Africa",165430.60,  183270.80,  198760.35,   214320.50, None, None],
        ["South Asia",        289640.25,  318450.70,  347830.60,   389210.40, None, None],
        ["Central Asia",       98720.40,  112340.55,  124670.80,   138950.20, None, None],
        ["Oceania",           178920.75,  192430.60,  207640.45,   231870.80, None, None],
        ["Canada",            312450.30,  328760.45,  367490.20,   412380.60, None, None],
        ["Japan",             467830.55,  489240.80,  512670.35,   546920.75, None, None],
        ["TOTAL",             None,       None,       None,        None,      None, None],
    ]

    # Number format for currency cells
    currency_fmt = '#,##0.00'
    percent_fmt = '0.00%'

    alt_fill = PatternFill(start_color="FFF0F4F8", end_color="FFF0F4F8", fill_type="solid")

    for r_idx, row_data in enumerate(data, 4):
        is_total = (row_data[0] == "TOTAL")
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if c_idx == 1:
                cell.value = val
                cell.font = Font(name="Calibri", size=11, bold=is_total)
                if not is_total and (r_idx % 2 == 0):
                    cell.fill = alt_fill
            else:
                cell.value = val
                if val is not None:
                    cell.number_format = currency_fmt if c_idx <= 6 else percent_fmt
                cell.font = Font(name="Calibri", size=11, bold=is_total)
                if not is_total and (r_idx % 2 == 0):
                    cell.fill = alt_fill
            cell.border = border

        if is_total:
            # Bold total row styling
            total_fill = PatternFill(start_color="FFCFD8DC", end_color="FFCFD8DC", fill_type="solid")
            for c_idx in range(1, 8):
                ws.cell(row=r_idx, column=c_idx).fill = total_fill

    # Add SUM formulas for columns B-F in total row (row 16)
    total_row = 16
    for col_letter, col_num in [("B", 2), ("C", 3), ("D", 4), ("E", 5)]:
        cell = ws.cell(row=total_row, column=col_num)
        cell.value = f"=SUM({col_letter}4:{col_letter}15)"
        cell.number_format = currency_fmt
        cell.font = Font(name="Calibri", size=11, bold=True)

    # Column F = sum of all quarters for each region (Annual Total column, rows 4-15)
    for r_idx in range(4, 16):
        cell = ws.cell(row=r_idx, column=6)
        cell.value = f"=B{r_idx}+C{r_idx}+D{r_idx}+E{r_idx}"
        cell.number_format = currency_fmt
        if r_idx == 15:
            cell.value = f"=SUM(F4:F15)"

    # Column G = YoY change as percentage (Q4 vs Q1 growth)
    for r_idx in range(4, 16):
        cell = ws.cell(row=r_idx, column=7)
        cell.value = f"=(E{r_idx}-B{r_idx})/B{r_idx}"
        cell.number_format = percent_fmt

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18

    # Freeze panes below header row
    ws.freeze_panes = "A4"

    # --- Sheet 2: Notes ---
    ws2 = wb.create_sheet("Notes")
    ws2["A1"] = "Report Notes"
    ws2["A1"].font = Font(name="Calibri", size=13, bold=True)
    ws2["A2"] = "Prepared by:"
    ws2["B2"] = "Finance Department"
    ws2["A3"] = "Date:"
    ws2["B3"] = "2025-03-01"
    ws2["A4"] = "Currency:"
    ws2["B4"] = "USD"
    ws2["A5"] = "Decimal Format:"
    ws2["B5"] = "Standard (period as decimal separator)"
    ws2["A6"] = "Note:"
    ws2["B6"] = "This report will be distributed in multiple regional formats"
    ws2["A7"] = "Status:"
    ws2["B7"] = "Awaiting decimal separator configuration"

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 45

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
