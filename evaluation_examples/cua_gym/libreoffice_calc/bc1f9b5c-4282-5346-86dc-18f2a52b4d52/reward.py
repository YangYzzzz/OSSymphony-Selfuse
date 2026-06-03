"""
Reward Script: Financial Report Decimal Separator Documentation
Task ID: osworld_calc_decimal_separator_005
Domain: libreoffice_calc
Scoring:
  Component 1 (0.6): Cell A1 in 'Financial Report' sheet contains 'Decimal Separator: Period (.)'
  Component 2 (0.4): Notes sheet B7 status updated to indicate decimal separator configured
Total: 1.0

The task requires the agent to:
1. Save a copy of the report with comma decimal separators (for regional audiences)
2. Switch back to period decimal separators for the English version
3. Document the final setting by writing 'Decimal Separator: Period (.)' in cell A1
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_decimal_separator_005'

def verify_task(file_path):
    """
    Verify task completion with progressive scoring.

    The task is complete when:
    1. Cell A1 of 'Financial Report' sheet contains 'Decimal Separator: Period (.)'
    2. Notes sheet B7 reflects that decimal separator has been configured (not 'Awaiting')

    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Cell A1 in 'Financial Report' sheet contains the required documentation text (0.6 points)
    # This is the primary task requirement: document the final decimal separator setting in A1
    # Initial state has 'Meridian Capital — Regional Financial Report 2025' in A1
    # Golden state requires 'Decimal Separator: Period (.)' in A1
    try:
        if 'Financial Report' not in wb.sheetnames:
            print("FAIL: Component 1 — Sheet 'Financial Report' not found")
        else:
            ws = wb['Financial Report']
            a1_value = ws['A1'].value
            if a1_value is None:
                print(f"FAIL: Component 1 — A1 is empty, expected 'Decimal Separator: Period (.)'")
            else:
                a1_str = str(a1_value).strip()
                # Check for exact match (case-insensitive, allowing minor whitespace variation)
                if 'Decimal Separator: Period (.)' in a1_str or a1_str.lower() == 'decimal separator: period (.)':
                    print(f"PASS: Component 1 — A1 contains required text: '{a1_str}' (0.6 pts)")
                    total_score += 0.6
                else:
                    print(f"FAIL: Component 1 — A1 = '{a1_str}', expected text containing 'Decimal Separator: Period (.)'")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Notes sheet B7 status updated to confirm decimal separator is configured (0.4 points)
    # Initial state: B7 = 'Awaiting decimal separator configuration'
    # Golden state: B7 should NOT say 'Awaiting' and should indicate configuration is done
    try:
        if 'Notes' not in wb.sheetnames:
            print("FAIL: Component 2 — Sheet 'Notes' not found")
        else:
            ws_notes = wb['Notes']
            b7_value = ws_notes['B7'].value
            if b7_value is None:
                print(f"FAIL: Component 2 — Notes!B7 is empty")
            else:
                b7_str = str(b7_value).strip()
                # Initial state has 'Awaiting decimal separator configuration'
                # Golden state should reflect it's been configured (not 'Awaiting')
                if 'Awaiting' in b7_str:
                    print(f"FAIL: Component 2 — Notes!B7 still says 'Awaiting': '{b7_str}'")
                elif 'period' in b7_str.lower() or 'configured' in b7_str.lower() or 'decimal' in b7_str.lower():
                    print(f"PASS: Component 2 — Notes!B7 updated to configured state: '{b7_str}' (0.4 pts)")
                    total_score += 0.4
                else:
                    print(f"FAIL: Component 2 — Notes!B7 = '{b7_str}', does not indicate decimal separator configured")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
