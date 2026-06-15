"""
Initial Setup: Create data_a.xlsx and data_b.ods on the Desktop with single columns of numbers,
and open a terminal window in the Desktop directory.
Task ID: osworld_multi_apps_terminal_calc_001
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user/Desktop'
TASK_ID = 'osworld_multi_apps_terminal_calc_001'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_data_a_xlsx():
    """Create data_a.xlsx with a single column of 12 numeric values."""
    output_path = f'{WORKDIR}/data_a.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Realistic numeric data — monthly revenue figures (in thousands)
    data_a_values = [
        142.5,
        238.0,
        195.75,
        307.25,
        412.0,
        289.50,
        358.00,
        475.25,
        521.50,
        399.75,
        463.00,
        512.25,
    ]

    ws.cell(row=1, column=1, value="Value")
    for r, val in enumerate(data_a_values, 2):
        ws.cell(row=r, column=1, value=val)

    wb.save(output_path)
    print(f'Created: {output_path}')
    return data_a_values


def create_data_b_ods():
    """Create data_b.ods with a single column of 10 numeric values using LibreOffice via macro."""
    output_path = f'{WORKDIR}/data_b.ods'

    # We create data_b.ods by first creating an xlsx, then converting it to ods via LibreOffice
    temp_xlsx = f'{WORKDIR}/data_b_temp.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Realistic numeric data — quarterly expense figures (in thousands)
    data_b_values = [
        88.40,
        117.60,
        205.30,
        156.80,
        93.20,
        178.50,
        243.70,
        132.45,
        189.90,
        211.15,
    ]

    ws.cell(row=1, column=1, value="Value")
    for r, val in enumerate(data_b_values, 2):
        ws.cell(row=r, column=1, value=val)

    wb.save(temp_xlsx)
    print(f'Created temp xlsx: {temp_xlsx}')

    # Convert temp xlsx to ods using LibreOffice headless
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    result = subprocess.run(
        [
            'libreoffice', '--headless', '--convert-to', 'ods',
            '--outdir', WORKDIR,
            temp_xlsx
        ],
        env=env,
        capture_output=True,
        text=True,
        timeout=60,
    )
    print(f'LibreOffice convert stdout: {result.stdout}')
    print(f'LibreOffice convert stderr: {result.stderr}')

    # LibreOffice names it data_b_temp.ods, rename to data_b.ods
    temp_ods = f'{WORKDIR}/data_b_temp.ods'
    if os.path.exists(temp_ods):
        os.rename(temp_ods, output_path)
        print(f'Renamed {temp_ods} -> {output_path}')
    else:
        print(f'WARNING: {temp_ods} not found, conversion may have failed')

    # Remove the temp xlsx
    if os.path.exists(temp_xlsx):
        os.remove(temp_xlsx)
        print(f'Removed temp: {temp_xlsx}')

    print(f'Created: {output_path}')
    return data_b_values


def create_initial():
    os.makedirs(WORKDIR, exist_ok=True)

    # Remove any pre-existing combined.csv so initial state is clean
    combined_path = f'{WORKDIR}/combined.csv'
    if os.path.exists(combined_path):
        os.remove(combined_path)
        print(f'Removed pre-existing: {combined_path}')

    # Create the two input files
    create_data_a_xlsx()
    create_data_b_ods()

    print('Both input files created on Desktop.')

    # GUI-ready startup: open a terminal window in the Desktop directory
    # Try gnome-terminal first, fall back to xterm
    try:
        launch_gui(
            f'gnome-terminal --working-directory="{WORKDIR}"',
            delay_sec=2.0,
        )
        print('GUI_READY: launched gnome-terminal in Desktop directory with DISPLAY=:0')
    except Exception as e:
        print(f'gnome-terminal failed ({e}), trying xterm...')
        launch_gui(f'xterm -e "cd {WORKDIR} && bash"', delay_sec=2.0)
        print('GUI_READY: launched xterm in Desktop directory with DISPLAY=:0')


create_initial()
