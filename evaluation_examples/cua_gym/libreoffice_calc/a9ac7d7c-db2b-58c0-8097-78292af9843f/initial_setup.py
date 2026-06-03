"""
Initial Setup: Set font to Courier New for response code cells
Task ID: calc_gfl_073
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_073'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Endpoints ---
    ws = wb.active
    ws.title = 'Endpoints'

    # Headers
    headers = ['Method', 'Path', 'Parameters', 'Response Code', 'Description', 'Authentication']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = header_align

    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 45
    ws.column_dimensions['F'].width = 18

    # 29 rows of realistic API endpoint data
    data = [
        ['GET', '/api/v2/users', 'page, limit, sort', '200 OK', 'Retrieve paginated list of all registered users', 'Bearer Token'],
        ['POST', '/api/v2/users', 'name, email, role', '201 Created', 'Create a new user account with specified role', 'Bearer Token'],
        ['GET', '/api/v2/users/{id}', 'id (path)', '200 OK', 'Retrieve detailed profile for a specific user', 'Bearer Token'],
        ['PUT', '/api/v2/users/{id}', 'id (path), name, email', '200 OK', 'Update all fields of an existing user record', 'Bearer Token'],
        ['DELETE', '/api/v2/users/{id}', 'id (path)', '204 No Content', 'Permanently remove a user and associated data', 'Admin Token'],
        ['GET', '/api/v2/products', 'category, min_price, max_price', '200 OK', 'Search and filter product catalog entries', 'API Key'],
        ['POST', '/api/v2/products', 'name, sku, price, stock', '201 Created', 'Add a new product to the inventory catalog', 'Admin Token'],
        ['GET', '/api/v2/products/{sku}', 'sku (path)', '200 OK', 'Retrieve product details by SKU identifier', 'API Key'],
        ['PATCH', '/api/v2/products/{sku}/stock', 'sku (path), quantity', '200 OK', 'Adjust stock quantity for a specific product', 'Admin Token'],
        ['DELETE', '/api/v2/products/{sku}', 'sku (path)', '404 Not Found', 'Remove a product listing from the catalog', 'Admin Token'],
        ['GET', '/api/v2/orders', 'status, date_from, date_to', '200 OK', 'List orders with optional status and date filters', 'Bearer Token'],
        ['POST', '/api/v2/orders', 'items[], shipping_address', '201 Created', 'Place a new order with line items and shipping info', 'Bearer Token'],
        ['GET', '/api/v2/orders/{id}', 'id (path)', '200 OK', 'Get full details of a specific order including items', 'Bearer Token'],
        ['PUT', '/api/v2/orders/{id}/status', 'id (path), status', '200 OK', 'Update order fulfillment status', 'Admin Token'],
        ['POST', '/api/v2/orders/{id}/cancel', 'id (path), reason', '200 OK', 'Cancel an existing order with cancellation reason', 'Bearer Token'],
        ['GET', '/api/v2/analytics/revenue', 'period, group_by', '200 OK', 'Aggregate revenue metrics grouped by time period', 'Admin Token'],
        ['GET', '/api/v2/analytics/traffic', 'date_range, source', '200 OK', 'Website traffic analytics with source attribution', 'Admin Token'],
        ['GET', '/api/v2/inventory/alerts', 'threshold', '200 OK', 'List products below minimum stock threshold', 'API Key'],
        ['POST', '/api/v2/auth/login', 'email, password', '200 OK', 'Authenticate user and return access token pair', 'None'],
        ['POST', '/api/v2/auth/refresh', 'refresh_token', '200 OK', 'Generate new access token from refresh token', 'None'],
        ['POST', '/api/v2/auth/logout', 'None', '204 No Content', 'Invalidate current session and revoke tokens', 'Bearer Token'],
        ['GET', '/api/v2/categories', 'parent_id', '200 OK', 'List product categories with optional parent filter', 'API Key'],
        ['POST', '/api/v2/categories', 'name, parent_id, description', '201 Created', 'Create a new product category in the hierarchy', 'Admin Token'],
        ['GET', '/api/v2/reviews/{product_id}', 'product_id (path), rating', '200 OK', 'Retrieve customer reviews for a specific product', 'API Key'],
        ['POST', '/api/v2/reviews', 'product_id, rating, comment', '201 Created', 'Submit a new product review with rating', 'Bearer Token'],
        ['GET', '/api/v2/shipping/rates', 'origin, destination, weight', '200 OK', 'Calculate shipping rates for a given route', 'API Key'],
        ['POST', '/api/v2/webhooks', 'url, events[], secret', '201 Created', 'Register a new webhook endpoint for event notifications', 'Admin Token'],
        ['GET', '/api/v2/webhooks', 'active', '200 OK', 'List all registered webhook configurations', 'Admin Token'],
        ['DELETE', '/api/v2/webhooks/{id}', 'id (path)', '500 Internal Server Error', 'Remove a webhook registration (known intermittent failure)', 'Admin Token'],
    ]

    # Default data font - explicitly Calibri (NOT Courier New)
    data_font = Font(name='Calibri', size=11)

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
