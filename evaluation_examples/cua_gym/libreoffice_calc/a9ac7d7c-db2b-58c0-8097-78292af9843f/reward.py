"""
Reward Script: Set font to Courier New for cells D2:D30
Task ID: calc_gfl_073
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): Majority (>=80%) of D2:D30 have Courier New font
  Component 2 (0.3): ALL D2:D30 have Courier New font (completeness)
  Component 3 (0.2): Other columns retain original font (not changed to Courier New)
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_073'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Endpoints' sheet must exist
    if 'Endpoints' not in wb.sheetnames:
        print(f"CRITICAL: 'Endpoints' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Endpoints']

    # Component 1: Majority (>=80%) of D2:D30 have Courier New font (0.5 points)
    try:
        courier_count = 0
        total_cells = 0
        for row in range(2, 31):
            cell = ws.cell(row=row, column=4)
            total_cells += 1
            font_name = cell.font.name
            if font_name and font_name.lower() == 'courier new':
                courier_count += 1

        ratio = courier_count / total_cells if total_cells > 0 else 0
        if ratio >= 0.8:
            print(f"PASS: Component 1 — {courier_count}/{total_cells} cells in D2:D30 have Courier New ({ratio:.0%}) (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — Only {courier_count}/{total_cells} cells in D2:D30 have Courier New ({ratio:.0%}), need >=80%")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: ALL D2:D30 cells have Courier New font (0.3 points)
    try:
        all_courier = True
        fail_cells = []
        for row in range(2, 31):
            cell = ws.cell(row=row, column=4)
            font_name = cell.font.name
            if not font_name or font_name.lower() != 'courier new':
                all_courier = False
                fail_cells.append(f"D{row}={font_name}")

        if all_courier:
            print(f"PASS: Component 2 — All 29 cells D2:D30 have Courier New (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — Not all cells have Courier New. Exceptions: {fail_cells[:5]}{'...' if len(fail_cells) > 5 else ''}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: D2:D30 have Courier New AND other columns retain original font (0.2 points)
    # This compound check ensures the font change was targeted only at column D.
    # Anchored to the task change: only awards if D cells ARE Courier New AND others are NOT.
    try:
        # Sub-check A: at least 80% of D2:D30 are Courier New (anchors to task change)
        d_courier = sum(1 for r in range(2, 31)
                        if (ws.cell(row=r, column=4).font.name or '').lower() == 'courier new')
        d_ok = d_courier >= 23  # >=80% of 29

        # Sub-check B: other columns NOT changed to Courier New
        other_cols_ok = True
        changed_cells = []
        check_cols = [1, 2, 3, 5, 6]  # A, B, C, E, F
        check_rows = [2, 5, 10, 15, 20, 25, 30]
        for row in check_rows:
            for col in check_cols:
                cell = ws.cell(row=row, column=col)
                font_name = cell.font.name
                if font_name and font_name.lower() == 'courier new':
                    from openpyxl.utils import get_column_letter
                    coord = f"{get_column_letter(col)}{row}"
                    other_cols_ok = False
                    changed_cells.append(coord)

        if d_ok and other_cols_ok:
            print(f"PASS: Component 3 — D cells are Courier New AND other columns retain original font (0.2 pts)")
            total_score += 0.2
        elif not d_ok:
            print(f"FAIL: Component 3 — D column not yet Courier New ({d_courier}/29), cannot award targeted-change points")
        else:
            print(f"FAIL: Component 3 — Other columns changed to Courier New: {changed_cells[:5]}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
