"""
Initial Setup: Freeze first two columns on Employee List sheet
Task ID: calc_ps_076
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_076'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Employee List'

    # --- Headers (A through M) ---
    headers = [
        'ID', 'Name', 'Department', 'Salary', 'Start Date',
        'Email', 'Phone', 'Location', 'Title', 'Manager',
        'Performance Rating', 'Bonus', 'Status'
    ]
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Realistic Employee Data (200 rows) ---
    first_names = [
        'Sarah', 'Marcus', 'Elena', 'James', 'Priya', 'David', 'Mei', 'Carlos',
        'Amara', 'Robert', 'Yuki', 'Thomas', 'Fatima', 'Andre', 'Lena',
        'William', 'Zara', 'Michael', 'Anya', 'Benjamin', 'Olivia', 'Nathan',
        'Sofia', 'Derek', 'Ingrid', 'Paul', 'Reiko', 'Vincent', 'Clara', 'Omar'
    ]
    last_names = [
        'Chen', 'Johnson', 'Petrov', 'Williams', 'Sharma', 'Kim', 'Rodriguez',
        'Okafor', 'Smith', 'Tanaka', 'Brown', 'Al-Rashid', 'Weber', 'Johansson',
        'Davis', 'Nakamura', 'Fernandez', 'Lee', 'Kowalski', 'Andersen',
        'Martinez', 'Nguyen', 'Dubois', 'Patel', 'Murphy', 'Suzuki', 'Torres',
        'Lindqvist', 'Costa', 'Park'
    ]
    departments = [
        'Engineering', 'Marketing', 'Finance', 'Human Resources', 'Sales',
        'Operations', 'Legal', 'Product', 'Customer Support', 'Research'
    ]
    locations = [
        'New York', 'San Francisco', 'Chicago', 'Austin', 'Seattle',
        'Boston', 'Denver', 'Portland', 'Atlanta', 'Miami'
    ]
    titles = [
        'Software Engineer', 'Marketing Specialist', 'Financial Analyst',
        'HR Coordinator', 'Sales Representative', 'Operations Manager',
        'Legal Counsel', 'Product Manager', 'Support Specialist',
        'Research Scientist', 'Senior Developer', 'Data Analyst',
        'Account Executive', 'UX Designer', 'Project Manager',
        'DevOps Engineer', 'Content Strategist', 'Business Analyst',
        'Technical Writer', 'Quality Assurance Engineer'
    ]
    statuses = ['Active', 'Active', 'Active', 'Active', 'Active',
                'Active', 'Active', 'On Leave', 'Remote', 'Active']

    random.seed(42)

    for row_idx in range(2, 202):  # rows 2-201 = 200 data rows
        emp_id = f'EMP-{1000 + row_idx - 1:04d}'
        first = random.choice(first_names)
        last = random.choice(last_names)
        name = f'{first} {last}'
        dept = random.choice(departments)
        salary = random.randint(55000, 165000)
        year = random.randint(2015, 2025)
        month = random.randint(1, 12)
        day = random.randint(1, 28)
        start_date = f'{year}-{month:02d}-{day:02d}'
        email = f'{first.lower()}.{last.lower()}@company.com'
        phone = f'({random.randint(200,999)}) {random.randint(100,999)}-{random.randint(1000,9999)}'
        location = random.choice(locations)
        title = random.choice(titles)
        manager_first = random.choice(first_names)
        manager_last = random.choice(last_names)
        manager = f'{manager_first} {manager_last}'
        perf = round(random.uniform(2.5, 5.0), 1)
        bonus = round(salary * random.uniform(0.03, 0.15), 2)
        status = random.choice(statuses)

        row_data = [emp_id, name, dept, salary, start_date, email, phone,
                    location, title, manager, perf, bonus, status]
        for col, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=val)

    # --- Column widths for readability ---
    col_widths = {
        'A': 12, 'B': 22, 'C': 18, 'D': 14, 'E': 14,
        'F': 30, 'G': 18, 'H': 16, 'I': 24, 'J': 22,
        'K': 20, 'L': 14, 'M': 12
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Row 1 height
    ws.row_dimensions[1].height = 20

    # NO freeze panes - this is the task for the agent to do
    ws.freeze_panes = None

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
