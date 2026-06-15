"""
Reward Script: Freeze first two columns (A and B) on 'Employee List' sheet
Task ID: calc_ps_076
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Freeze panes is set (not None)
  Component 2 (0.6): Freeze panes is exactly 'C1' (correct position)
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_076'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Employee List' sheet must exist
    if 'Employee List' not in wb.sheetnames:
        print(f"CRITICAL: Sheet 'Employee List' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Employee List']

    # Component 1: Freeze panes is set (not None) (0.4 points)
    # This checks that some freeze has been applied — fails on initial (None)
    try:
        freeze_val = ws.freeze_panes
        if freeze_val is not None:
            print(f"PASS: Component 1 — Freeze panes is set: '{freeze_val}' (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — Freeze panes is None (no freeze applied)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Freeze panes is exactly 'C1' (0.6 points)
    # Freezing at C1 means columns A and B are frozen, which is the task requirement.
    # This fails on initial (None) and only passes when correctly set to C1.
    try:
        freeze_val = ws.freeze_panes
        if freeze_val is not None and str(freeze_val) == 'C1':
            print(f"PASS: Component 2 — Freeze panes is 'C1' (columns A-B frozen) (0.6 pts)")
            total_score += 0.6
        else:
            print(f"FAIL: Component 2 — Expected freeze_panes='C1', found: '{freeze_val}'")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist app state before verification
persist_app_state("libreoffice_calc")

# Test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
