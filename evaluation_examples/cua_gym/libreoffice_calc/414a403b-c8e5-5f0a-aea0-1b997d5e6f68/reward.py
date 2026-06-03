"""
Reward Script: Insert 'Pivot Summary' sheet after 'Raw Data' with 'Summary Table' in A1
Task ID: calc_gg1_018
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): 'Pivot Summary' sheet exists
  Component 2 (0.3): Sheet tab order is correct (Raw Data, Pivot Summary, Charts, Notes)
  Component 3 (0.2): Cell A1 on 'Pivot Summary' contains 'Summary Table'
  Component 4 (0.2): 'Pivot Summary' sheet is otherwise empty and other sheets unaffected
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_018'


def persist_app_state(domain):
    """Save any unsaved LibreOffice state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    sheet_names = wb.sheetnames
    print(f"INFO: Found sheets: {sheet_names}")

    # Component 1: 'Pivot Summary' sheet exists (0.3 points)
    try:
        if 'Pivot Summary' in sheet_names:
            print(f"PASS: Component 1 - 'Pivot Summary' sheet exists (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 - 'Pivot Summary' sheet not found in {sheet_names}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Sheet tab order is correct (0.3 points)
    # Expected: Raw Data, Pivot Summary, Charts, Notes
    # Key requirement: Pivot Summary at index 1 (right after Raw Data, before Charts)
    try:
        expected_order = ['Raw Data', 'Pivot Summary', 'Charts', 'Notes']
        if sheet_names == expected_order:
            print(f"PASS: Component 2 - Sheet order is correct: {sheet_names} (0.3 pts)")
            total_score += 0.3
        else:
            # Partial: check if Pivot Summary is at least between Raw Data and Charts
            if 'Pivot Summary' in sheet_names:
                ps_idx = sheet_names.index('Pivot Summary')
                rd_idx = sheet_names.index('Raw Data') if 'Raw Data' in sheet_names else -1
                ch_idx = sheet_names.index('Charts') if 'Charts' in sheet_names else len(sheet_names)
                if rd_idx >= 0 and rd_idx < ps_idx < ch_idx:
                    print(f"PARTIAL: Component 2 - 'Pivot Summary' is between 'Raw Data' and 'Charts' but full order differs: {sheet_names} (0.15 pts)")
                    total_score += 0.15
                else:
                    print(f"FAIL: Component 2 - 'Pivot Summary' not in correct position. Order: {sheet_names}")
            else:
                print(f"FAIL: Component 2 - 'Pivot Summary' sheet missing, cannot check order")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Cell A1 on 'Pivot Summary' contains 'Summary Table' (0.2 points)
    try:
        if 'Pivot Summary' in sheet_names:
            ws_pivot = wb['Pivot Summary']
            a1_val = ws_pivot.cell(row=1, column=1).value
            if a1_val is not None and str(a1_val).strip() == 'Summary Table':
                print(f"PASS: Component 3 - A1 contains 'Summary Table' (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 3 - A1 expected 'Summary Table', found: {a1_val!r}")
        else:
            print(f"FAIL: Component 3 - 'Pivot Summary' sheet missing")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: 'Pivot Summary' is otherwise empty AND other sheets unaffected (0.2 points)
    try:
        if 'Pivot Summary' in sheet_names:
            ws_pivot = wb['Pivot Summary']
            # Check that the sheet only has content in A1 (max_row=1, max_col=1)
            is_empty_except_a1 = (ws_pivot.max_row <= 1 and ws_pivot.max_column <= 1)

            # Check other sheets still exist and have expected content
            other_sheet_issues = []
            # Raw Data should have header 'Region' in A1
            if 'Raw Data' in sheet_names:
                rd_a1 = wb['Raw Data'].cell(row=1, column=1).value
                if rd_a1 != 'Region':
                    other_sheet_issues.append(f"Raw Data A1 changed from 'Region' to {rd_a1!r}")
            else:
                other_sheet_issues.append("Raw Data sheet missing")

            # Charts should have 'Regional Sales Analysis' in A1
            if 'Charts' in sheet_names:
                ch_a1 = wb['Charts'].cell(row=1, column=1).value
                if ch_a1 != 'Regional Sales Analysis':
                    other_sheet_issues.append(f"Charts A1 changed from 'Regional Sales Analysis' to {ch_a1!r}")
            else:
                other_sheet_issues.append("Charts sheet missing")

            # Notes should have 'Sales Data Notes' in A1
            if 'Notes' in sheet_names:
                nt_a1 = wb['Notes'].cell(row=1, column=1).value
                if nt_a1 != 'Sales Data Notes':
                    other_sheet_issues.append(f"Notes A1 changed from 'Sales Data Notes' to {nt_a1!r}")
            else:
                other_sheet_issues.append("Notes sheet missing")

            other_sheets_ok = len(other_sheet_issues) == 0
            for issue in other_sheet_issues:
                print(f"  INFO: {issue}")

            if is_empty_except_a1 and other_sheets_ok:
                print(f"PASS: Component 4 - 'Pivot Summary' is empty except A1 and other sheets are intact (0.2 pts)")
                total_score += 0.2
            elif is_empty_except_a1:
                print(f"PARTIAL: Component 4 - 'Pivot Summary' is empty except A1 but some other sheets changed (0.1 pts)")
                total_score += 0.1
            elif other_sheets_ok:
                print(f"PARTIAL: Component 4 - Other sheets OK but 'Pivot Summary' has extra content (max_row={ws_pivot.max_row}, max_col={ws_pivot.max_column}) (0.1 pts)")
                total_score += 0.1
            else:
                print(f"FAIL: Component 4 - 'Pivot Summary' has extra content and other sheets changed")
        else:
            print(f"FAIL: Component 4 - 'Pivot Summary' sheet missing")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
