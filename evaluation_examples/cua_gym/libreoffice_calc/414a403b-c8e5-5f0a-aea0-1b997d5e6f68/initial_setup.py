"""
Initial Setup: Insert a 'Pivot Summary' sheet after 'Raw Data' with 'Summary Table' in A1
Task ID: calc_gg1_018
Domain: libreoffice_calc

Creates a workbook with three sheets: Raw Data, Charts, Notes.
The workbook contains realistic sales data for a data analyst scenario.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_018'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Raw Data ---
    ws1 = wb.active
    ws1.title = 'Raw Data'

    headers = ['Region', 'Product', 'Q1 Sales', 'Q2 Sales', 'Q3 Sales', 'Q4 Sales', 'Total']
    header_font = Font(bold=True, size=11, name='Calibri')
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center')
    white_font = Font(bold=True, size=11, name='Calibri', color='FFFFFF')

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    data = [
        ['North America', 'Widget Pro', 45230, 51280, 48750, 62340, None],
        ['North America', 'GadgetX', 32100, 28950, 35600, 41200, None],
        ['Europe', 'Widget Pro', 38470, 42100, 39800, 45670, None],
        ['Europe', 'GadgetX', 27850, 31200, 29450, 33800, None],
        ['Europe', 'SensorMax', 15320, 17800, 16450, 19200, None],
        ['Asia Pacific', 'Widget Pro', 52600, 58900, 55300, 67800, None],
        ['Asia Pacific', 'GadgetX', 41200, 44500, 42800, 49300, None],
        ['Asia Pacific', 'SensorMax', 23400, 26700, 24900, 29100, None],
        ['Latin America', 'Widget Pro', 18750, 21300, 19800, 24500, None],
        ['Latin America', 'GadgetX', 12400, 14200, 13100, 16800, None],
        ['Latin America', 'SensorMax', 8900, 10500, 9700, 12300, None],
        ['Middle East', 'Widget Pro', 14200, 16800, 15300, 19400, None],
        ['Middle East', 'GadgetX', 9800, 11400, 10600, 13700, None],
        ['Middle East', 'SensorMax', 6200, 7500, 6800, 8900, None],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c)
            if c == 7:
                # Total column: formula summing Q1-Q4
                cell.value = f'=SUM(C{r}:F{r})'
            elif val is not None:
                cell.value = val
            if c >= 3 and c <= 7:
                cell.number_format = '#,##0'

    # Set column widths
    ws1.column_dimensions['A'].width = 18
    ws1.column_dimensions['B'].width = 14
    for col_letter in ['C', 'D', 'E', 'F', 'G']:
        ws1.column_dimensions[col_letter].width = 12

    ws1.freeze_panes = 'A2'

    # --- Sheet 2: Charts ---
    ws2 = wb.create_sheet('Charts')

    ws2['A1'] = 'Regional Sales Analysis'
    ws2['A1'].font = Font(bold=True, size=14, name='Calibri')

    ws2['A3'] = 'Chart Data Summary'
    ws2['A3'].font = Font(bold=True, size=11, name='Calibri')

    chart_headers = ['Region', 'Total Revenue', '% of Global']
    for col, h in enumerate(chart_headers, 1):
        cell = ws2.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True)

    chart_data = [
        ['North America', 345450, '22.3%'],
        ['Europe', 311110, '20.1%'],
        ['Asia Pacific', 516500, '33.4%'],
        ['Latin America', 181950, '11.8%'],
        ['Middle East', 190600, '12.3%'],
    ]
    for r, row_data in enumerate(chart_data, 5):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 14

    # --- Sheet 3: Notes ---
    ws3 = wb.create_sheet('Notes')

    ws3['A1'] = 'Sales Data Notes'
    ws3['A1'].font = Font(bold=True, size=14, name='Calibri')

    notes = [
        ['Date', 'Author', 'Note'],
        ['2025-03-15', 'Sarah Chen', 'Q1 data finalized for all regions. Widget Pro remains top seller.'],
        ['2025-04-02', 'Marcus Johnson', 'Europe numbers adjusted after currency conversion correction.'],
        ['2025-04-10', 'Sarah Chen', 'Asia Pacific showing strongest growth trajectory at 33.4% share.'],
        ['2025-04-18', 'David Park', 'Latin America SensorMax underperforming - need marketing review.'],
        ['2025-05-01', 'Marcus Johnson', 'Q2 preliminary numbers added. Awaiting Middle East confirmation.'],
        ['2025-05-12', 'Sarah Chen', 'All Q2 figures confirmed. Ready for quarterly review presentation.'],
        ['2025-06-01', 'David Park', 'Q3 data collection started. New tracking template in use.'],
    ]

    for r, row_data in enumerate(notes, 1):
        for c, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            if r == 1:
                # This is actually row 2 visually since A1 has title
                pass

    # Re-do: put title in A1, headers in A3, data from A4
    ws3.delete_rows(1, ws3.max_row)
    ws3['A1'] = 'Sales Data Notes'
    ws3['A1'].font = Font(bold=True, size=14, name='Calibri')
    ws3['A3'] = 'Date'
    ws3['B3'] = 'Author'
    ws3['C3'] = 'Note'
    for cell in [ws3['A3'], ws3['B3'], ws3['C3']]:
        cell.font = Font(bold=True)

    notes_data = [
        ['2025-03-15', 'Sarah Chen', 'Q1 data finalized for all regions. Widget Pro remains top seller.'],
        ['2025-04-02', 'Marcus Johnson', 'Europe numbers adjusted after currency conversion correction.'],
        ['2025-04-10', 'Sarah Chen', 'Asia Pacific showing strongest growth trajectory at 33.4% share.'],
        ['2025-04-18', 'David Park', 'Latin America SensorMax underperforming - need marketing review.'],
        ['2025-05-01', 'Marcus Johnson', 'Q2 preliminary numbers added. Awaiting Middle East confirmation.'],
        ['2025-05-12', 'Sarah Chen', 'All Q2 figures confirmed. Ready for quarterly review presentation.'],
        ['2025-06-01', 'David Park', 'Q3 data collection started. New tracking template in use.'],
    ]
    for r, row_data in enumerate(notes_data, 4):
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r, column=c, value=val)

    ws3.column_dimensions['A'].width = 14
    ws3.column_dimensions['B'].width = 18
    ws3.column_dimensions['C'].width = 60

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheets: {wb.sheetnames}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
