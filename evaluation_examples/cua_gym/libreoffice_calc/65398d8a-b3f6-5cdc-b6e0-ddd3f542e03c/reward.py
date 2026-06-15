"""
Reward Script: Generate a pivot table showing count of tasks by status for each project
Task ID: calc_pivot_030
Domain: libreoffice_calc
Scoring:
  Component 1 (0.15): PivotTable sheet exists
  Component 2 (0.25): Correct structure — Project rows, Status columns, header labels
  Component 3 (0.35): Correct count values for key cells
  Component 4 (0.25): Grand Total row and column are correct
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_030'


def persist_app_state(domain: str):
    """Try to save any unsaved GUI state via Ctrl+S."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: PivotTable sheet exists (0.15 points)
    # This is the core task-introduced change: initial has only 'ProjectTasks'
    pivot_ws = None
    try:
        non_source_sheets = [sn for sn in wb.sheetnames if sn != 'ProjectTasks']
        if len(non_source_sheets) > 0:
            pivot_ws = wb[non_source_sheets[0]]
            print(f"PASS: Component 1 — Found pivot sheet '{non_source_sheets[0]}' (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — No pivot table sheet found. Sheets: {wb.sheetnames}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    if pivot_ws is None:
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Read the pivot table data into a structured dict for flexible verification
    # The pivot table should have: Project as rows, Status categories as columns
    try:
        # Read header row
        headers = []
        for c in range(1, pivot_ws.max_column + 1):
            val = pivot_ws.cell(row=1, column=c).value
            if val is not None:
                headers.append(str(val).strip())
            else:
                headers.append(None)

        # Build a dict: {project_name: {status_label: count}}
        pivot_data = {}
        project_col = 0  # first column is always the row label (Project)
        for r in range(2, pivot_ws.max_row + 1):
            row_label = pivot_ws.cell(row=r, column=1).value
            if row_label is None:
                continue
            row_label = str(row_label).strip()
            row_dict = {}
            for c in range(2, len(headers) + 1):
                col_header = headers[c - 1] if c - 1 < len(headers) else None
                if col_header is not None:
                    cell_val = pivot_ws.cell(row=r, column=c).value
                    row_dict[col_header] = cell_val
            pivot_data[row_label] = row_dict

        print(f"  Pivot headers: {headers}")
        print(f"  Pivot rows: {list(pivot_data.keys())}")
    except Exception as e:
        print(f"ERROR: Could not parse pivot table: {e}")
        pivot_data = {}
        headers = []

    # Component 2: Correct structure — Project rows and Status columns (0.25 points)
    # Must have: 4 project rows (Alpha, Beta, Gamma, Delta) and status column headers
    try:
        expected_projects = {'Alpha', 'Beta', 'Gamma', 'Delta'}
        expected_statuses = {'Not Started', 'In Progress', 'Review', 'Done'}

        # Check projects exist as row labels
        found_projects = set()
        for key in pivot_data.keys():
            if key in expected_projects:
                found_projects.add(key)

        # Check status categories appear somewhere in headers
        # Headers may be like "Count of Not Started" or just "Not Started"
        found_statuses = set()
        for h in headers[1:]:  # skip first (Project label)
            if h is None:
                continue
            for status in expected_statuses:
                if status.lower() in h.lower():
                    found_statuses.add(status)

        has_projects = found_projects == expected_projects
        has_statuses = found_statuses == expected_statuses

        if has_projects and has_statuses:
            print(f"PASS: Component 2 — All 4 projects and 4 status columns found (0.25 pts)")
            total_score += 0.25
        elif has_projects or has_statuses:
            partial = 0.12
            print(f"PARTIAL: Component 2 — Projects: {has_projects}, Statuses: {has_statuses} ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Projects found: {found_projects}, Statuses found: {found_statuses}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Correct count values for key cells (0.35 points)
    # Ground truth: Alpha/Done=18, Beta/In Progress=15, Grand total=180
    # Also check a few more: Alpha total=45, Beta total=45
    try:
        checks_passed = 0
        total_checks = 5

        # Helper to find a value in pivot_data by project and status substring
        def find_value(project, status_substr):
            if project not in pivot_data:
                return None
            for key, val in pivot_data[project].items():
                if key is not None and status_substr.lower() in key.lower():
                    return val
            return None

        # Check 1: Alpha / Done = 18
        val = find_value('Alpha', 'Done')
        if val is not None and int(val) == 18:
            checks_passed += 1
            print(f"  CHECK: Alpha/Done = {val} == 18 PASS")
        else:
            print(f"  CHECK: Alpha/Done = {val}, expected 18 FAIL")

        # Check 2: Beta / In Progress = 15
        val = find_value('Beta', 'In Progress')
        if val is not None and int(val) == 15:
            checks_passed += 1
            print(f"  CHECK: Beta/In Progress = {val} == 15 PASS")
        else:
            print(f"  CHECK: Beta/In Progress = {val}, expected 15 FAIL")

        # Check 3: Gamma / Not Started = 12
        val = find_value('Gamma', 'Not Started')
        if val is not None and int(val) == 12:
            checks_passed += 1
            print(f"  CHECK: Gamma/Not Started = {val} == 12 PASS")
        else:
            print(f"  CHECK: Gamma/Not Started = {val}, expected 12 FAIL")

        # Check 4: Delta / Review = 13
        val = find_value('Delta', 'Review')
        if val is not None and int(val) == 13:
            checks_passed += 1
            print(f"  CHECK: Delta/Review = {val} == 13 PASS")
        else:
            print(f"  CHECK: Delta/Review = {val}, expected 13 FAIL")

        # Check 5: Alpha / In Progress = 10
        val = find_value('Alpha', 'In Progress')
        if val is not None and int(val) == 10:
            checks_passed += 1
            print(f"  CHECK: Alpha/In Progress = {val} == 10 PASS")
        else:
            print(f"  CHECK: Alpha/In Progress = {val}, expected 10 FAIL")

        comp3_score = 0.35 * (checks_passed / total_checks)
        if checks_passed > 0:
            total_score += comp3_score
            if checks_passed == total_checks:
                print(f"PASS: Component 3 — All {total_checks} count checks passed ({comp3_score:.2f} pts)")
            else:
                print(f"PARTIAL: Component 3 — {checks_passed}/{total_checks} count checks passed ({comp3_score:.2f} pts)")
        else:
            print(f"FAIL: Component 3 — 0/{total_checks} count checks passed")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Grand Total row and column correctness (0.25 points)
    # Grand Total row should exist and show total = 180
    try:
        grand_total_row = None
        for key in pivot_data.keys():
            if 'grand' in key.lower() or 'total' in key.lower():
                grand_total_row = pivot_data[key]
                break

        if grand_total_row is None:
            print(f"FAIL: Component 4 — No Grand Total row found. Row labels: {list(pivot_data.keys())}")
        else:
            # Check that Grand Total column (last column or 'Grand Total' key) = 180
            gt_value = None
            for key, val in grand_total_row.items():
                if key is not None and 'grand' in key.lower() or key is not None and 'total' in key.lower():
                    gt_value = val
            # Also check last value in the row
            if gt_value is None:
                # Try last value
                vals = list(grand_total_row.values())
                if vals:
                    gt_value = vals[-1]

            sub_checks = 0
            # Sub-check 1: Grand Total exists
            sub_checks += 1  # already found the row

            # Sub-check 2: Grand Total value = 180
            if gt_value is not None and int(gt_value) == 180:
                sub_checks += 1
                print(f"  CHECK: Grand Total = {gt_value} == 180 PASS")
            else:
                print(f"  CHECK: Grand Total = {gt_value}, expected 180 FAIL")

            # Sub-check 3: Column totals sum correctly (Not Started=40, In Progress=45, Review=40, Done=55)
            col_totals_correct = 0
            expected_col_totals = {'Not Started': 40, 'In Progress': 45, 'Review': 40, 'Done': 55}
            for status, expected_val in expected_col_totals.items():
                for key, val in grand_total_row.items():
                    if key is not None and status.lower() in key.lower():
                        if val is not None and int(val) == expected_val:
                            col_totals_correct += 1
                        break

            if col_totals_correct == 4:
                sub_checks += 1
                print(f"  CHECK: All column totals correct PASS")
            else:
                print(f"  CHECK: {col_totals_correct}/4 column totals correct FAIL")

            comp4_score = 0.25 * (sub_checks / 3)
            if sub_checks > 0:
                total_score += comp4_score
                if sub_checks == 3:
                    print(f"PASS: Component 4 — Grand Total row fully correct ({comp4_score:.2f} pts)")
                else:
                    print(f"PARTIAL: Component 4 — {sub_checks}/3 sub-checks passed ({comp4_score:.2f} pts)")
            else:
                print(f"FAIL: Component 4 — 0/3 sub-checks passed")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
