"""
Initial Setup: Generate a pivot table from project tracking sheet
Task ID: calc_pivot_030
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_030'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ProjectTasks'

    # --- Headers ---
    headers = ['TaskID', 'Project', 'Assignee', 'Status', 'Priority', 'Hours']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Data distribution ---
    # 4 projects, 45 tasks each = 180 total
    # Statuses: Not Started, In Progress, Review, Done
    # Constraint: Alpha/Done=18, Beta/In Progress=15, Grand total=180
    projects = ['Alpha', 'Beta', 'Gamma', 'Delta']
    statuses = ['Not Started', 'In Progress', 'Review', 'Done']

    # Distribution: dict of (project, status) -> count
    # Each project gets 45 tasks
    distribution = {
        ('Alpha', 'Not Started'): 8,
        ('Alpha', 'In Progress'): 10,
        ('Alpha', 'Review'): 9,
        ('Alpha', 'Done'): 18,  # ground truth
        ('Beta', 'Not Started'): 10,
        ('Beta', 'In Progress'): 15,  # ground truth
        ('Beta', 'Review'): 8,
        ('Beta', 'Done'): 12,
        ('Gamma', 'Not Started'): 12,
        ('Gamma', 'In Progress'): 11,
        ('Gamma', 'Review'): 10,
        ('Gamma', 'Done'): 12,
        ('Delta', 'Not Started'): 10,
        ('Delta', 'In Progress'): 9,
        ('Delta', 'Review'): 13,
        ('Delta', 'Done'): 13,
    }

    # Verify totals
    total = sum(distribution.values())
    assert total == 180, f"Total is {total}, expected 180"
    for p in projects:
        proj_total = sum(v for (pp, s), v in distribution.items() if pp == p)
        assert proj_total == 45, f"{p} total is {proj_total}, expected 45"

    # Build task list
    assignees = [
        'Sarah Chen', 'Marcus Johnson', 'Emily Rodriguez', 'David Kim',
        'Lisa Patel', 'James Wilson', 'Maria Garcia', 'Robert Taylor',
        'Anna Lee', 'Michael Brown', 'Jennifer White', 'Thomas Anderson',
        'Rachel Green', 'Daniel Martinez', 'Sophie Turner', 'Chris Evans',
        'Priya Sharma', 'Alex Nguyen', 'Olivia Scott', 'Kevin O\'Brien'
    ]
    priorities = ['Low', 'Medium', 'High']

    tasks = []
    for (project, status), count in distribution.items():
        for _ in range(count):
            tasks.append((project, status))

    # Shuffle to make it look realistic (not grouped)
    random.shuffle(tasks)

    # Write rows
    for i, (project, status) in enumerate(tasks):
        row = i + 2
        task_id = f'T{i+1:03d}'
        assignee = random.choice(assignees)
        priority = random.choice(priorities)
        hours = random.randint(2, 40)

        ws.cell(row=row, column=1, value=task_id)
        ws.cell(row=row, column=2, value=project)
        ws.cell(row=row, column=3, value=assignee)
        ws.cell(row=row, column=4, value=status)
        ws.cell(row=row, column=5, value=priority)
        ws.cell(row=row, column=6, value=hours)

    # --- Column widths ---
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 8

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
