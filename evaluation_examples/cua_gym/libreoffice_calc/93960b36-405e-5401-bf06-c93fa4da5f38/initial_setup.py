"""
Initial Setup: Mixed-case and whitespace data causing SUMIF mismatch
Task ID: calc_tbl_075
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_075'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: SalesData ---
    ws = wb.active
    ws.title = 'SalesData'

    # Headers
    headers = ['Region', 'Product', 'Amount', 'Quarter', 'Rep']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows - intentionally messy Region column with mixed case and whitespace
    data = [
        ['East',   'Widget A',    4520.00, 'Q1', 'Sarah Chen'],
        ['West',   'Widget B',    3890.50, 'Q1', 'Marcus Johnson'],
        ['east',   'Gadget Pro',  6210.75, 'Q1', 'Priya Patel'],
        ['North',  'Widget A',    2980.00, 'Q2', 'James Wilson'],
        ['EAST',   'Gadget Pro',  5175.25, 'Q2', 'Aisha Rahman'],
        ['South',  'Widget B',    4100.00, 'Q2', 'Carlos Mendez'],
        ['East ',  'Widget A',    3845.50, 'Q2', 'Sarah Chen'],
        ['West',   'Gadget Pro',  7320.00, 'Q3', 'Marcus Johnson'],
        [' East',  'Widget B',    2960.00, 'Q3', 'Priya Patel'],
        ['North',  'Widget A',    4450.25, 'Q3', 'James Wilson'],
        ['South',  'Gadget Pro',  3675.00, 'Q3', 'Aisha Rahman'],
        ['West',   'Widget A',    5890.75, 'Q4', 'Carlos Mendez'],
        ['North',  'Widget B',    3210.00, 'Q4', 'James Wilson'],
        ['South',  'Widget A',    4780.50, 'Q4', 'Carlos Mendez'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 3:  # Amount column
                cell.number_format = '$#,##0.00'

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 18

    # --- Summary section below data ---
    summary_row = 18
    ws.cell(row=summary_row, column=1, value='Summary').font = Font(bold=True, size=12)

    ws.cell(row=summary_row + 1, column=1, value='Region')
    ws.cell(row=summary_row + 1, column=2, value='Total Sales')
    ws.cell(row=summary_row + 1, column=1).font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=2).font = Font(bold=True)

    # SUMIF formulas - these only match exact "East", missing 'east', 'EAST', 'East ', ' East'
    regions = ['East', 'West', 'North', 'South']
    for i, region in enumerate(regions):
        row = summary_row + 2 + i
        ws.cell(row=row, column=1, value=region)
        ws.cell(row=row, column=2, value=f'=SUMIF(A2:A15,"{region}",C2:C15)')
        ws.cell(row=row, column=2).number_format = '$#,##0.00'

    # Note about the problem
    note_row = summary_row + 7
    ws.cell(row=note_row, column=1, value='NOTE: East total seems too low - check for data entry inconsistencies')
    ws.cell(row=note_row, column=1).font = Font(italic=True, color='FF0000')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
