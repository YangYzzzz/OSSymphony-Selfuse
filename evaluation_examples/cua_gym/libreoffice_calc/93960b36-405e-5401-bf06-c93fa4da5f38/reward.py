"""
Reward Script: Clean mixed-case and whitespace region data for SUMIF
Task ID: calc_tbl_075
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4 pts): Case variants ('east', 'EAST') cleaned to 'East'
  Component 2 (0.3 pts): Whitespace variants ('East ', ' East') cleaned to 'East'
  Component 3 (0.3 pts): All 5 East-region entries are uniformly 'East'
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_075'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice document."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Initial state has these Region values in A2:A15:
      A2='East', A4='east', A6='EAST', A8='East ' (trailing space), A10=' East' (leading space)
      Others: A3='West', A5='North', A7='South', A9='West', A11='North', A12='South',
              A13='West', A14='North', A15='South'

    Golden state: A4, A6, A8, A10 all cleaned to 'East' (no extra whitespace, consistent case).
    The SUMIF formula in B20 should now match all 5 East entries.
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    try:
        ws = wb['SalesData']
    except KeyError:
        print("CRITICAL: Sheet 'SalesData' not found")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Case variants cleaned (0.4 points)
    # A4 was 'east', A6 was 'EAST' — both should now be exactly 'East'
    try:
        a4_val = ws['A4'].value
        a6_val = ws['A6'].value
        case_fixed_count = 0
        if isinstance(a4_val, str) and a4_val == 'East':
            case_fixed_count += 1
            print(f"  PASS: A4 is 'East' (was 'east')")
        else:
            print(f"  FAIL: A4 expected 'East', found: {a4_val!r}")

        if isinstance(a6_val, str) and a6_val == 'East':
            case_fixed_count += 1
            print(f"  PASS: A6 is 'East' (was 'EAST')")
        else:
            print(f"  FAIL: A6 expected 'East', found: {a6_val!r}")

        if case_fixed_count == 2:
            print(f"PASS: Component 1 — Both case variants cleaned (0.4 pts)")
            total_score += 0.4
        elif case_fixed_count == 1:
            print(f"PARTIAL: Component 1 — 1 of 2 case variants cleaned (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 1 — No case variants cleaned")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Whitespace variants cleaned (0.3 points)
    # A8 was 'East ' (trailing space), A10 was ' East' (leading space)
    # Both should now be exactly 'East'
    try:
        a8_val = ws['A8'].value
        a10_val = ws['A10'].value
        ws_fixed_count = 0
        if isinstance(a8_val, str) and a8_val == 'East':
            ws_fixed_count += 1
            print(f"  PASS: A8 is 'East' (was 'East ')")
        else:
            print(f"  FAIL: A8 expected 'East', found: {a8_val!r}")

        if isinstance(a10_val, str) and a10_val == 'East':
            ws_fixed_count += 1
            print(f"  PASS: A10 is 'East' (was ' East')")
        else:
            print(f"  FAIL: A10 expected 'East', found: {a10_val!r}")

        if ws_fixed_count == 2:
            print(f"PASS: Component 2 — Both whitespace variants cleaned (0.3 pts)")
            total_score += 0.3
        elif ws_fixed_count == 1:
            print(f"PARTIAL: Component 2 — 1 of 2 whitespace variants cleaned (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 — No whitespace variants cleaned")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: All 5 East-region entries are uniformly 'East' (0.3 points)
    # East entries should be at rows 2, 4, 6, 8, 10 — all exactly 'East'
    try:
        east_rows = [2, 4, 6, 8, 10]
        exact_east_count = 0
        for r in east_rows:
            val = ws.cell(row=r, column=1).value
            if isinstance(val, str) and val == 'East':
                exact_east_count += 1

        if exact_east_count == 5:
            print(f"PASS: Component 3 — All 5 East entries are uniformly 'East' (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 3 — {exact_east_count}/5 entries are exactly 'East'")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
