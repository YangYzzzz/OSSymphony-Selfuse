"""
Initial Setup: Create procurement spreadsheet with 250 rows of supplier order data.
Task ID: calc_pivot_076
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_076'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

random.seed(42)


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


# 8 suppliers with target average delivery days (ascending)
# These define the mean for random generation
SUPPLIERS = {
    'Apex Materials Co.':        {'mean_days': 5,  'count': 30},
    'BlueRidge Supply Ltd.':     {'mean_days': 8,  'count': 32},
    'CedarPoint Logistics':      {'mean_days': 10, 'count': 35},
    'Dominion Parts Inc.':       {'mean_days': 12, 'count': 28},
    'EverGreen Components':      {'mean_days': 14, 'count': 33},
    'FairTrade Industrial':      {'mean_days': 17, 'count': 30},
    'GlobalTech Distributors':   {'mean_days': 19, 'count': 31},
    'Highland Raw Materials':    {'mean_days': 22, 'count': 31},
}

# Verify total = 250
assert sum(s['count'] for s in SUPPLIERS.values()) == 250


def generate_data():
    """Generate 250 rows of procurement data."""
    rows = []
    po_num = 1

    for supplier, info in SUPPLIERS.items():
        mean = info['mean_days']
        count = info['count']
        for _ in range(count):
            # Delivery days: normal dist around mean, clamp 1-30
            dd = max(1, min(30, int(random.gauss(mean, 3))))

            # Order date in 2025
            order_date = datetime(2025, 1, 1) + timedelta(days=random.randint(0, 300))
            delivery_date = order_date + timedelta(days=dd)

            # Order amount: $500 - $25,000
            amount = round(random.uniform(500, 25000), 2)

            rows.append({
                'poid': f'PO{po_num:03d}',
                'supplier': supplier,
                'order_date': order_date,
                'delivery_date': delivery_date,
                'delivery_days': dd,
                'order_amount': amount,
            })
            po_num += 1

    # Shuffle to make it realistic (not grouped by supplier)
    random.shuffle(rows)
    return rows


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Procurement'

    # --- Headers ---
    headers = ['POID', 'Supplier', 'OrderDate', 'DeliveryDate', 'DeliveryDays', 'OrderAmount']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Data ---
    data = generate_data()

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data['poid'])
        ws.cell(row=r, column=2, value=row_data['supplier'])
        ws.cell(row=r, column=3, value=row_data['order_date'])
        ws.cell(row=r, column=3).number_format = 'yyyy-mm-dd'
        ws.cell(row=r, column=4, value=row_data['delivery_date'])
        ws.cell(row=r, column=4).number_format = 'yyyy-mm-dd'
        ws.cell(row=r, column=5, value=row_data['delivery_days'])
        ws.cell(row=r, column=6, value=row_data['order_amount'])
        ws.cell(row=r, column=6).number_format = '$#,##0.00'

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 16

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
