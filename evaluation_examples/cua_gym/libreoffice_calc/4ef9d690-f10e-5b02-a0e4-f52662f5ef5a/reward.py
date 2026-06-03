"""
Reward Script: Create pivot table showing supplier performance
Task ID: calc_pivot_076
Domain: libreoffice_calc
Scoring:
  Component 1: Pivot Table sheet exists with headers (0.15)
  Component 2: Correct supplier names and avg delivery days (0.35)
  Component 3: Correct order counts per supplier (0.20)
  Component 4: Sorted ascending by avg delivery days (0.15)
  Component 5: Grand Total row with correct values (0.15)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_076'

# Expected data derived from task description and source data
EXPECTED_SUPPLIERS = [
    ("Apex Materials Co.", 4.33, 30),
    ("BlueRidge Supply Ltd.", 9.31, 32),
    ("CedarPoint Logistics", 9.46, 35),
    ("Dominion Parts Inc.", 12.29, 28),
    ("EverGreen Components", 14.39, 33),
    ("FairTrade Industrial", 17.17, 30),
    ("GlobalTech Distributors", 19.1, 31),
    ("Highland Raw Materials", 21.06, 31),
]
EXPECTED_GRAND_TOTAL_AVG = 13.35
EXPECTED_GRAND_TOTAL_COUNT = 250


def persist_app_state(domain):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def find_pivot_sheet(wb):
    """Find the pivot table sheet (not 'Procurement')."""
    for name in wb.sheetnames:
        if name.lower() != 'procurement':
            return wb[name], name
    return None, None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Pivot Table sheet exists with appropriate headers (0.15 points)
    try:
        pivot_ws, pivot_name = find_pivot_sheet(wb)
        if pivot_ws is None:
            print("FAIL: Component 1 — No pivot table sheet found (only 'Procurement' exists)")
            print("REWARD: 0.0")
            return 0.0

        print(f"INFO: Found pivot sheet: '{pivot_name}'")

        # Check that there are at least header row + 8 data rows
        if pivot_ws.max_row < 9:
            print(f"FAIL: Component 1 — Pivot sheet has only {pivot_ws.max_row} rows, expected at least 10")
        else:
            # Check for header-like content: must have columns for supplier, avg days, count
            header_found = False
            for row_num in range(1, min(4, pivot_ws.max_row + 1)):
                row_vals = []
                for col in range(1, pivot_ws.max_column + 1):
                    v = pivot_ws.cell(row=row_num, column=col).value
                    if v is not None:
                        row_vals.append(str(v).lower())
                row_text = ' '.join(row_vals)
                if ('supplier' in row_text or 'name' in row_text) and ('delivery' in row_text or 'days' in row_text or 'avg' in row_text or 'count' in row_text or 'order' in row_text):
                    header_found = True
                    break
            if header_found:
                print(f"PASS: Component 1 — Pivot sheet '{pivot_name}' exists with headers (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 1 — Pivot sheet exists but headers not recognized")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    if pivot_ws is None:
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Locate data: find the header row and column mapping
    header_row = None
    col_supplier = None
    col_avg = None
    col_count = None

    try:
        for row_num in range(1, min(5, pivot_ws.max_row + 1)):
            for col in range(1, pivot_ws.max_column + 1):
                v = pivot_ws.cell(row=row_num, column=col).value
                if v is not None:
                    vl = str(v).lower()
                    if 'supplier' in vl or vl == 'name':
                        col_supplier = col
                        header_row = row_num
                    elif 'delivery' in vl or 'avg' in vl or 'average' in vl:
                        col_avg = col
                    elif 'count' in vl or 'order count' in vl:
                        col_count = col
    except Exception as e:
        print(f"ERROR: Header detection — {e}")

    if header_row is None or col_supplier is None:
        print("FAIL: Could not locate supplier column in pivot table")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    print(f"INFO: Header row={header_row}, supplier_col={col_supplier}, avg_col={col_avg}, count_col={col_count}")

    # Read all data rows (skip header, skip grand total)
    data_rows = []
    grand_total_row = None
    for row_num in range(header_row + 1, pivot_ws.max_row + 1):
        supplier_val = pivot_ws.cell(row=row_num, column=col_supplier).value
        if supplier_val is None:
            continue
        supplier_str = str(supplier_val).strip()
        if supplier_str.lower() in ('grand total', 'total', 'sum', 'overall'):
            avg_val = pivot_ws.cell(row=row_num, column=col_avg).value if col_avg else None
            count_val = pivot_ws.cell(row=row_num, column=col_count).value if col_count else None
            grand_total_row = (supplier_str, avg_val, count_val)
            continue
        avg_val = pivot_ws.cell(row=row_num, column=col_avg).value if col_avg else None
        count_val = pivot_ws.cell(row=row_num, column=col_count).value if col_count else None
        data_rows.append((supplier_str, avg_val, count_val))

    print(f"INFO: Found {len(data_rows)} data rows, grand_total={'yes' if grand_total_row else 'no'}")

    # Component 2: Correct supplier names and avg delivery days (0.35 points)
    try:
        if len(data_rows) < 8:
            print(f"FAIL: Component 2 — Only {len(data_rows)} suppliers found, expected 8")
        else:
            matches = 0
            for expected_name, expected_avg, _ in EXPECTED_SUPPLIERS:
                for actual_name, actual_avg, _ in data_rows:
                    if actual_name.lower() == expected_name.lower():
                        if actual_avg is not None:
                            try:
                                if abs(float(actual_avg) - expected_avg) <= 0.1:
                                    matches += 1
                                    break
                            except (ValueError, TypeError):
                                pass
                        break

            if matches == 8:
                print(f"PASS: Component 2 — All 8 suppliers with correct avg delivery days (0.35 pts)")
                total_score += 0.35
            elif matches >= 6:
                partial = round(0.35 * (matches / 8), 2)
                print(f"PARTIAL: Component 2 — {matches}/8 suppliers correct ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 2 — Only {matches}/8 suppliers have correct avg delivery days")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Correct order counts per supplier (0.20 points)
    try:
        if col_count is None:
            print("FAIL: Component 3 — No order count column found")
        elif len(data_rows) < 8:
            print(f"FAIL: Component 3 — Only {len(data_rows)} supplier rows, expected 8")
        else:
            count_matches = 0
            for expected_name, _, expected_count in EXPECTED_SUPPLIERS:
                for actual_name, _, actual_count in data_rows:
                    if actual_name.lower() == expected_name.lower():
                        if actual_count is not None:
                            try:
                                if int(float(actual_count)) == expected_count:
                                    count_matches += 1
                            except (ValueError, TypeError):
                                pass
                        break

            if count_matches == 8:
                print(f"PASS: Component 3 — All 8 supplier order counts correct (0.20 pts)")
                total_score += 0.20
            elif count_matches >= 6:
                partial = round(0.20 * (count_matches / 8), 2)
                print(f"PARTIAL: Component 3 — {count_matches}/8 counts correct ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 3 — Only {count_matches}/8 supplier counts correct")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Sorted ascending by avg delivery days (0.15 points)
    try:
        if col_avg is None:
            print("FAIL: Component 4 — No avg delivery days column found")
        elif len(data_rows) < 2:
            print("FAIL: Component 4 — Not enough rows to verify sort order")
        else:
            avg_values = []
            for _, avg_val, _ in data_rows:
                if avg_val is not None:
                    try:
                        avg_values.append(float(avg_val))
                    except (ValueError, TypeError):
                        pass

            if len(avg_values) >= 2:
                is_sorted = all(avg_values[i] <= avg_values[i + 1] for i in range(len(avg_values) - 1))
                if is_sorted:
                    print(f"PASS: Component 4 — Sorted ascending by avg delivery days (0.15 pts)")
                    total_score += 0.15
                else:
                    print(f"FAIL: Component 4 — Not sorted ascending. Values: {avg_values}")
            else:
                print(f"FAIL: Component 4 — Could not extract enough avg values to check sort")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Grand Total row with correct values (0.15 points)
    try:
        if grand_total_row is None:
            print("FAIL: Component 5 — No Grand Total row found")
        else:
            gt_name, gt_avg, gt_count = grand_total_row
            gt_pass = True

            # Check grand total count = 250
            if gt_count is not None:
                try:
                    if int(float(gt_count)) != EXPECTED_GRAND_TOTAL_COUNT:
                        print(f"FAIL: Component 5 — Grand total count={gt_count}, expected {EXPECTED_GRAND_TOTAL_COUNT}")
                        gt_pass = False
                except (ValueError, TypeError):
                    print(f"FAIL: Component 5 — Grand total count not numeric: {gt_count}")
                    gt_pass = False
            else:
                print("FAIL: Component 5 — Grand total count is None")
                gt_pass = False

            # Check grand total avg ~ 13.35
            if gt_avg is not None:
                try:
                    if abs(float(gt_avg) - EXPECTED_GRAND_TOTAL_AVG) > 0.2:
                        print(f"FAIL: Component 5 — Grand total avg={gt_avg}, expected ~{EXPECTED_GRAND_TOTAL_AVG}")
                        gt_pass = False
                except (ValueError, TypeError):
                    print(f"FAIL: Component 5 — Grand total avg not numeric: {gt_avg}")
                    gt_pass = False
            else:
                print("FAIL: Component 5 — Grand total avg is None")
                gt_pass = False

            if gt_pass:
                print(f"PASS: Component 5 — Grand Total row correct: avg={gt_avg}, count={gt_count} (0.15 pts)")
                total_score += 0.15
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
persist_app_state("libreoffice_calc")
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
