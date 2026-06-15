"""
Reward Script: Use VLOOKUP to assign letter grades and GPA points to 45 students
Task ID: calc_edu_vlookup_gradesheet_006
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Column C (Letter Grade) — VLOOKUP formulas for all 45 rows    0.4 pts
  Component 2: Column D (GPA Points) — VLOOKUP formulas for all 45 rows       0.4 pts
  Component 3: Column D number format set to 1 decimal place ('0.0')           0.2 pts
  Total: 1.0

Task changes from initial → golden:
  - Rows 2-46, col C: None → VLOOKUP(B2,GradeScale.$A$2:$C$6,2,1)
  - Rows 2-46, col D: None → VLOOKUP(B2,GradeScale.$A$2:$C$6,3,1)
  - Col D number_format: General → 0.0
  - GradeScale sheet is unchanged (precondition, not scored)
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_edu_vlookup_gradesheet_006'

EXPECTED_LOOKUP_RANGE = 'GradeScale.$A$2:$C$6'
TOTAL_STUDENTS = 45  # rows 2 to 46


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Scores' sheet must exist
    if 'Scores' not in wb.sheetnames:
        print("FAIL: 'Scores' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Scores']

    # -----------------------------------------------------------------------
    # Component 1: Column C (Letter Grade) — VLOOKUP formulas for all 45 rows
    # (0.4 points)
    # Each row C2:C46 must contain a VLOOKUP formula that:
    #   - References the corresponding B cell as lookup value
    #   - Uses GradeScale.$A$2:$C$6 as the table array
    #   - Returns column index 2 (Letter Grade)
    #   - Uses approximate match (last arg = 1 or TRUE)
    # -----------------------------------------------------------------------
    try:
        c_vlookup_rows = 0
        c_wrong_rows = []

        for row in range(2, TOTAL_STUDENTS + 2):  # rows 2-46
            cell_val = ws.cell(row=row, column=3).value

            # Must be a formula string containing VLOOKUP
            if not isinstance(cell_val, str):
                c_wrong_rows.append((row, cell_val))
                continue

            formula_upper = cell_val.upper().replace(' ', '')

            # Check: VLOOKUP present
            if 'VLOOKUP' not in formula_upper:
                c_wrong_rows.append((row, cell_val))
                continue

            # Check: references the correct B cell (e.g., B2 for row 2)
            expected_b_ref = f'B{row}'
            if expected_b_ref not in cell_val and expected_b_ref.lower() not in cell_val.lower():
                c_wrong_rows.append((row, cell_val))
                continue

            # Check: references the GradeScale lookup table
            lookup_range_upper = EXPECTED_LOOKUP_RANGE.upper().replace(' ', '')
            if lookup_range_upper not in formula_upper:
                c_wrong_rows.append((row, cell_val))
                continue

            # Check: column index is 2 (Letter Grade)
            # Pattern: VLOOKUP(Bx,GradeScale.$A$2:$C$6,2,...)
            if ',2,' not in formula_upper and ',2)' not in formula_upper:
                c_wrong_rows.append((row, cell_val))
                continue

            c_vlookup_rows += 1

        if c_vlookup_rows == TOTAL_STUDENTS:
            print(f"PASS: Component 1 — All {TOTAL_STUDENTS} rows in Column C have correct VLOOKUP formula (0.4 pts)")
            total_score += 0.4
        elif c_vlookup_rows >= TOTAL_STUDENTS * 0.9:
            # Partial credit: at least 90% of rows correct
            partial = round(0.4 * (c_vlookup_rows / TOTAL_STUDENTS), 2)
            print(f"PARTIAL: Component 1 — {c_vlookup_rows}/{TOTAL_STUDENTS} rows in Column C have VLOOKUP; wrong rows: {c_wrong_rows[:5]}")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {c_vlookup_rows}/{TOTAL_STUDENTS} rows in Column C have VLOOKUP; wrong rows sample: {c_wrong_rows[:5]}")

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Column D (GPA Points) — VLOOKUP formulas for all 45 rows
    # (0.4 points)
    # Each row D2:D46 must contain a VLOOKUP formula that:
    #   - References the corresponding B cell as lookup value
    #   - Uses GradeScale.$A$2:$C$6 as the table array
    #   - Returns column index 3 (GPA)
    #   - Uses approximate match (last arg = 1 or TRUE)
    # -----------------------------------------------------------------------
    try:
        d_vlookup_rows = 0
        d_wrong_rows = []

        for row in range(2, TOTAL_STUDENTS + 2):  # rows 2-46
            cell_val = ws.cell(row=row, column=4).value

            if not isinstance(cell_val, str):
                d_wrong_rows.append((row, cell_val))
                continue

            formula_upper = cell_val.upper().replace(' ', '')

            if 'VLOOKUP' not in formula_upper:
                d_wrong_rows.append((row, cell_val))
                continue

            # Check: references the correct B cell
            expected_b_ref = f'B{row}'
            if expected_b_ref not in cell_val and expected_b_ref.lower() not in cell_val.lower():
                d_wrong_rows.append((row, cell_val))
                continue

            # Check: references the GradeScale lookup table
            lookup_range_upper = EXPECTED_LOOKUP_RANGE.upper().replace(' ', '')
            if lookup_range_upper not in formula_upper:
                d_wrong_rows.append((row, cell_val))
                continue

            # Check: column index is 3 (GPA Points)
            if ',3,' not in formula_upper and ',3)' not in formula_upper:
                d_wrong_rows.append((row, cell_val))
                continue

            d_vlookup_rows += 1

        if d_vlookup_rows == TOTAL_STUDENTS:
            print(f"PASS: Component 2 — All {TOTAL_STUDENTS} rows in Column D have correct VLOOKUP formula (0.4 pts)")
            total_score += 0.4
        elif d_vlookup_rows >= TOTAL_STUDENTS * 0.9:
            partial = round(0.4 * (d_vlookup_rows / TOTAL_STUDENTS), 2)
            print(f"PARTIAL: Component 2 — {d_vlookup_rows}/{TOTAL_STUDENTS} rows in Column D have VLOOKUP; wrong rows: {d_wrong_rows[:5]}")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {d_vlookup_rows}/{TOTAL_STUDENTS} rows in Column D have VLOOKUP; wrong rows sample: {d_wrong_rows[:5]}")

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Column D (GPA Points) — number format is 1 decimal place
    # (0.2 points)
    # At least 40 of the 45 D-column cells must have number_format = '0.0'
    # This is a change from the initial file where D column is 'General'.
    # -----------------------------------------------------------------------
    try:
        d_format_count = 0
        d_format_wrong = []

        for row in range(2, TOTAL_STUDENTS + 2):  # rows 2-46
            cell = ws.cell(row=row, column=4)
            nf = cell.number_format
            # Accept '0.0' format (1 decimal place)
            if nf == '0.0':
                d_format_count += 1
            else:
                d_format_wrong.append((row, nf))

        if d_format_count == TOTAL_STUDENTS:
            print(f"PASS: Component 3 — All {TOTAL_STUDENTS} GPA Points cells formatted to 1 decimal place '0.0' (0.2 pts)")
            total_score += 0.2
        elif d_format_count >= TOTAL_STUDENTS * 0.8:
            partial = round(0.2 * (d_format_count / TOTAL_STUDENTS), 2)
            print(f"PARTIAL: Component 3 — {d_format_count}/{TOTAL_STUDENTS} GPA Points cells formatted '0.0'; wrong sample: {d_format_wrong[:5]}")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {d_format_count}/{TOTAL_STUDENTS} GPA Points cells have '0.0' format; wrong sample: {d_format_wrong[:5]}")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
