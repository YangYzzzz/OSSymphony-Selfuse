"""
Initial Setup: VLOOKUP grade sheet — create Scores and GradeScale sheets
Task ID: calc_edu_vlookup_gradesheet_006
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_vlookup_gradesheet_006'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Scores ---
    ws_scores = wb.active
    ws_scores.title = 'Scores'

    # Headers
    ws_scores['A1'] = 'Student Name'
    ws_scores['B1'] = 'Final %'
    ws_scores['C1'] = 'Letter Grade'
    ws_scores['D1'] = 'GPA Points'

    # 45 students with realistic names and percentage scores
    students = [
        ('Aiden Brooks',      92.4),
        ('Brianna Torres',    78.1),
        ('Carlos Mendez',     65.7),
        ('Diana Nguyen',      88.3),
        ('Ethan Kowalski',    55.0),
        ('Fiona Patel',       91.6),
        ('Gabriel Chen',      74.2),
        ('Hannah Williams',   83.5),
        ('Isaac Okonkwo',     47.8),
        ('Julia Kim',         96.1),
        ('Kevin Ramirez',     62.9),
        ('Lauren Mitchell',   79.4),
        ('Marcus Thompson',   85.7),
        ('Natalie Russo',     71.3),
        ('Oliver Johansson',  90.0),
        ('Priya Sharma',      68.5),
        ('Quinn Hartley',     58.2),
        ('Rachel Goldstein',  87.9),
        ('Samuel Adeyemi',    73.6),
        ('Tiffany Larson',    94.3),
        ('Umar Saleh',        61.4),
        ('Victoria Huang',    80.8),
        ('William Foster',    52.1),
        ('Xena Petrov',       89.2),
        ('Yasmine El-Amin',   76.5),
        ('Zachary Morgan',    93.7),
        ('Alexis Reyes',      67.0),
        ('Blake Nakamura',    82.4),
        ('Chloe Andersson',   57.8),
        ('Derek Sullivan',    95.6),
        ('Elena Vasquez',     70.9),
        ('Finn O\'Brien',     84.1),
        ('Grace Liu',         48.3),
        ('Hugo Fernandez',    77.2),
        ('Isabelle Dumont',   91.0),
        ('Jaylen Washington', 63.5),
        ('Kathryn Barnes',    86.4),
        ('Liam Hoffmann',     72.8),
        ('Mia Castellano',    59.3),
        ('Nolan Eriksson',    98.2),
        ('Olivia Osei',       75.6),
        ('Patrick Donnelly',  81.7),
        ('Quinn Blackwood',   44.5),
        ('Rosa Delgado',      88.9),
        ('Stefan Korhonen',   69.4),
    ]

    for row_idx, (name, pct) in enumerate(students, 2):
        ws_scores.cell(row=row_idx, column=1, value=name)
        ws_scores.cell(row=row_idx, column=2, value=pct)
        # Columns C and D intentionally left empty (task asks to fill with VLOOKUP)

    # --- Sheet 2: GradeScale ---
    ws_grade = wb.create_sheet('GradeScale')

    # Headers
    ws_grade['A1'] = 'Min%'
    ws_grade['B1'] = 'Letter'
    ws_grade['C1'] = 'GPA'

    # Grade scale rows (sorted ascending — required for approximate VLOOKUP)
    grade_data = [
        (0,  'F', 0),
        (60, 'D', 1),
        (70, 'C', 2),
        (80, 'B', 3),
        (90, 'A', 4),
    ]
    for row_idx, (min_pct, letter, gpa) in enumerate(grade_data, 2):
        ws_grade.cell(row=row_idx, column=1, value=min_pct)
        ws_grade.cell(row=row_idx, column=2, value=letter)
        ws_grade.cell(row=row_idx, column=3, value=gpa)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: Scores (45 students, C/D empty), GradeScale (A1:C6)')


create_initial()
