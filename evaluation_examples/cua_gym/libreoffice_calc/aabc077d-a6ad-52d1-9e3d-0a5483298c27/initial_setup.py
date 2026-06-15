"""
Initial Setup: Create CustomerRevenue spreadsheet with monthly data (no charts)
Task ID: calc_chart_combo_area_line_049
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_combo_area_line_049'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: CustomerRevenue ---
    ws = wb.active
    ws.title = 'CustomerRevenue'

    # Headers (Row 1)
    ws['A1'] = 'Month'
    ws['B1'] = 'Revenue ($000)'
    ws['C1'] = 'Customers'

    # Style headers - bold with light blue background
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}1']
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows (Row 2-7) - exactly as specified in context
    data = [
        ('Jan', 245, 1820),
        ('Feb', 268, 1950),
        ('Mar', 312, 2140),
        ('Apr', 298, 2080),
        ('May', 345, 2350),
        ('Jun', 389, 2620),
    ]

    for r, (month, revenue, customers) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=month)
        ws.cell(row=r, column=2, value=revenue)
        ws.cell(row=r, column=3, value=customers)

    # Column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14

    # Set row height for header
    ws.row_dimensions[1].height = 20

    # NOTE: No charts in the initial file (task requires creating one)
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheet: CustomerRevenue')
    print('Data rows: 6 (Jan-Jun)')
    print('No charts present (as required by task)')


create_initial()
