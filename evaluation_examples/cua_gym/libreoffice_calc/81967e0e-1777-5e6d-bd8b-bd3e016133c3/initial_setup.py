"""
Initial Setup: Rename sheet, insert new sheet, change tab color
Task ID: calc_ggf_008
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_008'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = 'Summary'

    summary_headers = ['Metric', 'Q1 2024', 'Q2 2024', 'Q3 2024', 'Q4 2024']
    for col, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    summary_data = [
        ['Total Revenue', 245000, 312000, 287000, 356000],
        ['Operating Costs', 178000, 195000, 182000, 210000],
        ['Net Profit', 67000, 117000, 105000, 146000],
        ['New Customers', 142, 198, 167, 224],
        ['Customer Retention Rate', 0.89, 0.91, 0.87, 0.93],
        ['Avg Order Value', 385.50, 412.30, 397.20, 445.80],
        ['Employee Count', 45, 48, 52, 55],
        ['Support Tickets', 312, 287, 345, 298],
    ]
    for r, row_data in enumerate(summary_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_summary.cell(row=r, column=c, value=val)

    # Format retention rate as percentage
    for r in range(6, 7):
        for c in range(2, 6):
            ws_summary.cell(row=r, column=c).number_format = '0.00%'

    # Format currency columns
    for r in [2, 3, 4, 7]:
        for c in range(2, 6):
            ws_summary.cell(row=r, column=c).number_format = '$#,##0'

    ws_summary.column_dimensions['A'].width = 25
    for col_letter in ['B', 'C', 'D', 'E']:
        ws_summary.column_dimensions[col_letter].width = 15

    # --- Sheet 2: Data Import ---
    ws_data = wb.create_sheet('Data Import')

    data_headers = ['Record ID', 'Customer Name', 'Product', 'Quantity', 'Unit Price',
                    'Total Amount', 'Order Date', 'Region', 'Status']
    for col, h in enumerate(data_headers, 1):
        cell = ws_data.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    import_data = [
        ['IMP-2024-001', 'Meridian Technologies', 'Server Rack Unit', 3, 2450.00, 7350.00, '2024-01-12', 'Northeast', 'Delivered'],
        ['IMP-2024-002', 'Coastal Dynamics LLC', 'Network Switch 48-Port', 12, 875.00, 10500.00, '2024-01-18', 'Southeast', 'Delivered'],
        ['IMP-2024-003', 'Alpine Solutions Group', 'Firewall Appliance', 2, 3200.00, 6400.00, '2024-02-03', 'West', 'In Transit'],
        ['IMP-2024-004', 'Prairie Wind Energy', 'UPS Battery Backup', 8, 560.00, 4480.00, '2024-02-14', 'Midwest', 'Delivered'],
        ['IMP-2024-005', 'Summit Health Partners', 'SSD Storage 2TB', 25, 189.99, 4749.75, '2024-03-01', 'Northeast', 'Processing'],
        ['IMP-2024-006', 'Bayshore Logistics', 'Cable Management Kit', 50, 45.00, 2250.00, '2024-03-10', 'Southeast', 'Delivered'],
        ['IMP-2024-007', 'Redwood Analytics', 'Monitor 27" 4K', 15, 425.00, 6375.00, '2024-03-22', 'West', 'Delivered'],
        ['IMP-2024-008', 'Northstar Manufacturing', 'Keyboard Mechanical', 30, 129.99, 3899.70, '2024-04-05', 'Midwest', 'In Transit'],
        ['IMP-2024-009', 'Harbor View Financial', 'Docking Station USB-C', 20, 215.00, 4300.00, '2024-04-18', 'Northeast', 'Delivered'],
        ['IMP-2024-010', 'Cascade Software Inc', 'Webcam HD 1080p', 40, 79.99, 3199.60, '2024-05-02', 'West', 'Processing'],
        ['IMP-2024-011', 'Ironbridge Construction', 'Laptop Stand Adjustable', 18, 65.00, 1170.00, '2024-05-15', 'Southeast', 'Delivered'],
        ['IMP-2024-012', 'Silverline Consulting', 'Wireless Mouse Pro', 35, 54.99, 1924.65, '2024-06-01', 'Northeast', 'In Transit'],
    ]
    for r, row_data in enumerate(import_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_data.cell(row=r, column=c, value=val)

    # Format currency columns
    for r in range(2, len(import_data) + 2):
        ws_data.cell(row=r, column=5).number_format = '$#,##0.00'
        ws_data.cell(row=r, column=6).number_format = '$#,##0.00'

    ws_data.column_dimensions['A'].width = 18
    ws_data.column_dimensions['B'].width = 25
    ws_data.column_dimensions['C'].width = 25
    ws_data.column_dimensions['D'].width = 10
    ws_data.column_dimensions['E'].width = 12
    ws_data.column_dimensions['F'].width = 14
    ws_data.column_dimensions['G'].width = 14
    ws_data.column_dimensions['H'].width = 12
    ws_data.column_dimensions['I'].width = 12

    # --- Sheet 3: Charts ---
    ws_charts = wb.create_sheet('Charts')

    chart_headers = ['Region', 'Total Orders', 'Revenue', 'Avg Order Size']
    for col, h in enumerate(chart_headers, 1):
        cell = ws_charts.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    chart_data = [
        ['Northeast', 156, 482350.00, 3092.63],
        ['Southeast', 134, 398200.00, 2972.39],
        ['West', 189, 567800.00, 3004.23],
        ['Midwest', 121, 345600.00, 2856.20],
    ]
    for r, row_data in enumerate(chart_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_charts.cell(row=r, column=c, value=val)

    for r in range(2, 6):
        ws_charts.cell(row=r, column=3).number_format = '$#,##0.00'
        ws_charts.cell(row=r, column=4).number_format = '$#,##0.00'

    ws_charts.column_dimensions['A'].width = 15
    ws_charts.column_dimensions['B'].width = 14
    ws_charts.column_dimensions['C'].width = 14
    ws_charts.column_dimensions['D'].width = 16

    # Ensure no custom tab colors
    # (default is no tab color, which is what we want)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
