"""
Reward Script: Gantt-style project timeline with colored task duration bars
Task ID: calc_gpm_012
Domain: libreoffice_calc
Scoring:
  Component 1 (0.60) - Colored fills in correct cells matching task durations
  Component 2 (0.25) - Each task row uses a unique/distinct color
  Component 3 (0.15) - Consistency: each task uses the same color across its full span
NOTE: Thin borders in D2:K7 already exist in initial file (precondition), so NOT scored.
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_012'

# Task duration definitions: (row, start_week, end_week)
# Week columns: D=W1(col4), E=W2(col5), ..., K=W8(col11)
TASKS = [
    (2, 1, 2),   # Requirements: W1-W2
    (3, 2, 3),   # Design: W2-W3
    (4, 3, 6),   # Backend Dev: W3-W6
    (5, 4, 7),   # Frontend Dev: W4-W7
    (6, 6, 8),   # Testing: W6-W8
    (7, 8, 8),   # Deployment: W8
]

def week_to_col(week_num):
    """Convert week number (1-8) to column index (4-11, i.e. D-K)."""
    return week_num + 3  # W1=col4(D), W2=col5(E), ..., W8=col11(K)


def get_fill_color(cell):
    """Get the ARGB fill color of a cell, or None if no solid fill."""
    try:
        if cell.fill.fill_type == 'solid' and cell.fill.fgColor and cell.fill.fgColor.rgb:
            rgb = cell.fill.fgColor.rgb
            # Exclude default/no-color values
            if rgb not in ('00000000', '00000000'):
                return rgb
    except Exception:
        pass
    return None


def has_thin_border(cell):
    """Check if a cell has thin borders on all four sides."""
    try:
        sides = [cell.border.left, cell.border.right, cell.border.top, cell.border.bottom]
        return all(s is not None and s.style == 'thin' for s in sides)
    except Exception:
        return False


def verify_task(file_path):
    """
    Verify Gantt chart task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet 'Gantt' must exist
    if 'Gantt' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Gantt' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Gantt']

    # =========================================================================
    # Component 1: Colored fills in correct cells matching task durations (0.50)
    # Each task row should have colored cells spanning its duration weeks,
    # and NO colored cells outside its duration range.
    # This is the core Gantt chart feature and ONLY exists in golden, not initial.
    # =========================================================================
    try:
        correct_fill_count = 0
        total_fill_checks = 0
        wrong_fill_count = 0

        for (row, start_w, end_w) in TASKS:
            for col in range(4, 12):  # D(4) through K(11)
                week_num = col - 3  # col4=W1, col5=W2, ..., col11=W8
                cell = ws.cell(row=row, column=col)
                fill_color = get_fill_color(cell)

                if start_w <= week_num <= end_w:
                    # This cell SHOULD be colored
                    total_fill_checks += 1
                    if fill_color is not None:
                        correct_fill_count += 1
                    else:
                        print(f"  MISS: Row {row} col {col} (W{week_num}) should be colored but has no fill")
                else:
                    # This cell should NOT be colored
                    if fill_color is not None:
                        wrong_fill_count += 1
                        print(f"  EXTRA: Row {row} col {col} (W{week_num}) has fill {fill_color} but should be empty")

        # Calculate sub-score: ratio of correct fills, penalize wrong fills
        if total_fill_checks > 0:
            fill_ratio = correct_fill_count / total_fill_checks
            penalty = min(wrong_fill_count * 0.05, 0.25)  # cap penalty
            comp1_score = max(0.0, fill_ratio - penalty)
        else:
            comp1_score = 0.0

        comp1_points = round(0.60 * comp1_score, 4)
        if comp1_points > 0:
            total_score += comp1_points
        print(f"COMPONENT 1: Colored fills — {correct_fill_count}/{total_fill_checks} correct, "
              f"{wrong_fill_count} extra fills, sub-score={comp1_score:.2f} ({comp1_points} pts)")

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Each task row uses a unique/distinct color (0.25)
    # All 6 task rows must have different fill colors from each other.
    # This verifies the "color unique to that task" requirement.
    # =========================================================================
    try:
        row_colors = {}
        for (row, start_w, end_w) in TASKS:
            # Sample the first colored cell in the row's duration
            first_col = week_to_col(start_w)
            cell = ws.cell(row=row, column=first_col)
            fill_color = get_fill_color(cell)
            if fill_color is not None:
                row_colors[row] = fill_color

        # Check uniqueness: all colors must be distinct
        unique_colors = set(row_colors.values())
        tasks_with_color = len(row_colors)
        unique_count = len(unique_colors)

        if tasks_with_color == 6 and unique_count == 6:
            print(f"PASS: Component 2 — All 6 tasks have unique colors: {unique_colors}")
            total_score += 0.25
        elif tasks_with_color > 0 and unique_count == tasks_with_color:
            # Partial: some tasks colored but not all 6
            partial = 0.25 * (tasks_with_color / 6.0)
            total_score += round(partial, 4)
            print(f"PARTIAL: Component 2 — {tasks_with_color}/6 tasks have unique colors ({round(partial, 4)} pts)")
        elif tasks_with_color > 0:
            # Some tasks share colors
            ratio = unique_count / 6.0
            partial = 0.25 * ratio * 0.5  # heavy penalty for non-unique
            if partial > 0:
                total_score += round(partial, 4)
            print(f"PARTIAL: Component 2 — {unique_count} unique colors out of {tasks_with_color} colored tasks ({round(partial, 4)} pts)")
        else:
            print(f"FAIL: Component 2 — No tasks have colored fills")

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Color consistency within each task row (0.15)
    # Each task's filled cells should all use the SAME color across the span.
    # This checks that the Gantt bars are uniform, not random colors per cell.
    # Only scores if task has at least one colored cell (anchored to task change).
    # =========================================================================
    try:
        consistent_tasks = 0
        tasks_with_any_color = 0

        for (row, start_w, end_w) in TASKS:
            colors_in_span = []
            for week_num in range(start_w, end_w + 1):
                col = week_to_col(week_num)
                cell = ws.cell(row=row, column=col)
                fill_color = get_fill_color(cell)
                if fill_color is not None:
                    colors_in_span.append(fill_color)

            if len(colors_in_span) > 0:
                tasks_with_any_color += 1
                # All colors in the span should be the same
                if len(set(colors_in_span)) == 1 and len(colors_in_span) == (end_w - start_w + 1):
                    consistent_tasks += 1
                else:
                    print(f"  INCONSISTENT: Row {row} has colors {set(colors_in_span)}, "
                          f"filled {len(colors_in_span)}/{end_w - start_w + 1} cells")

        if tasks_with_any_color > 0:
            consistency_ratio = consistent_tasks / 6.0
            comp3_points = round(0.15 * consistency_ratio, 4)
            if comp3_points > 0:
                total_score += comp3_points
            if consistent_tasks == 6:
                print(f"PASS: Component 3 — All 6 tasks have consistent colors across their span ({comp3_points} pts)")
            else:
                print(f"PARTIAL: Component 3 — {consistent_tasks}/6 tasks consistent ({comp3_points} pts)")
        else:
            print(f"FAIL: Component 3 — No tasks have any colored fills to check consistency")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
