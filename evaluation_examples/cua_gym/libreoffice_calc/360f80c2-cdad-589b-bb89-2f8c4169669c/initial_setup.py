"""
Initial Setup: Create a Gantt-style project timeline spreadsheet
Task ID: calc_gpm_012
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_012'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Gantt'

    # --- Row 1: Headers ---
    headers = ['Task', 'Start Week', 'End Week',
               'W1', 'W2', 'W3', 'W4', 'W5', 'W6', 'W7', 'W8']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF505050", end_color="FF505050", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Rows 2-7: Task data ---
    tasks = [
        ['Requirements', 1, 2],
        ['Design', 2, 3],
        ['Backend Dev', 3, 6],
        ['Frontend Dev', 4, 7],
        ['Testing', 6, 8],
        ['Deployment', 8, 8],
    ]

    for r, (name, start, end) in enumerate(tasks, 2):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=start)
        ws.cell(row=r, column=3, value=end)

    # --- Thin borders on D2:K7 ---
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(2, 8):
        for col in range(4, 12):  # D=4, K=11
            ws.cell(row=row, column=col).border = border

    # --- Column widths D:K = 5 ---
    for col_letter in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        ws.column_dimensions[col_letter].width = 5

    # Reasonable widths for A, B, C
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
