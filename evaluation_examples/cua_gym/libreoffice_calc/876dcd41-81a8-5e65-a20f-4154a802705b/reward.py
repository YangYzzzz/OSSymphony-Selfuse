"""
Reward Script: Break circular reference in budget spreadsheet
Task ID: calc_tbl_013
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): E20 does NOT reference E25 (breaks the cycle)
  Component 2 (0.30): F25 contains a total/SUM formula
  Component 3 (0.20): E20 contains a partial sum formula (sums non-E20 rows)
  Component 4 (0.15): E25 still contains a SUM formula for the budget items
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_013'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get the Budget sheet
    try:
        ws = wb['Budget']
    except KeyError:
        # Try active sheet as fallback
        ws = wb.active
        if ws is None:
            print("CRITICAL: No Budget sheet and no active sheet")
            print("REWARD: 0.0")
            return 0.0
        print(f"WARNING: No 'Budget' sheet found, using active sheet '{ws.title}'")

    # Read formulas from key cells
    e20_val = ws['E20'].value
    e25_val = ws['E25'].value
    f25_val = ws['F25'].value

    print(f"DEBUG: E20 = {repr(e20_val)}")
    print(f"DEBUG: E25 = {repr(e25_val)}")
    print(f"DEBUG: F25 = {repr(f25_val)}")

    # Component 1: E20 does NOT reference E25 (0.35 points)
    # This is the core fix - breaking the circular dependency.
    # In the initial file, E20 = =E25-SUM(E1:E19) which references E25.
    # After fix, E20 should NOT contain any reference to E25.
    try:
        if e20_val is not None and isinstance(e20_val, str) and e20_val.startswith('='):
            formula_upper = e20_val.upper()
            # Check that E25 is not referenced in the formula
            if 'E25' not in formula_upper:
                print(f"PASS: Component 1 - E20 does not reference E25 (0.35 pts)")
                total_score += 0.35
            else:
                print(f"FAIL: Component 1 - E20 still references E25: {e20_val}")
        else:
            print(f"FAIL: Component 1 - E20 is not a formula: {repr(e20_val)}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: F25 contains a total/SUM formula (0.30 points)
    # The task says to put the total in F25. It should be a SUM of budget items.
    try:
        if f25_val is not None and isinstance(f25_val, str) and f25_val.startswith('='):
            formula_upper = f25_val.upper()
            if 'SUM' in formula_upper or 'E1' in formula_upper:
                print(f"PASS: Component 2 - F25 has a total formula: {f25_val} (0.30 pts)")
                total_score += 0.30
            else:
                print(f"FAIL: Component 2 - F25 has a formula but not a sum/total: {f25_val}")
        else:
            print(f"FAIL: Component 2 - F25 does not contain a formula: {repr(f25_val)}")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: E20 contains a partial sum (sums budget items excluding E20 itself) AND
    # does NOT reference E25 (0.20 points)
    # The golden fix uses =SUM(E1:E19)+SUM(E21:E24) to sum all items except E20 and E25.
    # We accept any formula that sums partial ranges of budget items without referencing E25.
    try:
        if e20_val is not None and isinstance(e20_val, str) and e20_val.startswith('='):
            formula_upper = e20_val.upper()
            # Must NOT reference E25 (anchored to the change) AND must contain SUM
            no_e25_ref = 'E25' not in formula_upper
            has_partial_sum = ('SUM' in formula_upper and
                              ('E1' in formula_upper or 'E2' in formula_upper or
                               'E19' in formula_upper or 'E21' in formula_upper or
                               'E24' in formula_upper))
            if no_e25_ref and has_partial_sum:
                print(f"PASS: Component 3 - E20 contains a partial sum formula without E25 ref: {e20_val} (0.20 pts)")
                total_score += 0.20
            elif not no_e25_ref:
                print(f"FAIL: Component 3 - E20 still references E25: {e20_val}")
            else:
                print(f"FAIL: Component 3 - E20 formula doesn't seem to be a partial sum: {e20_val}")
        else:
            print(f"FAIL: Component 3 - E20 is not a formula: {repr(e20_val)}")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: E25 still contains a SUM formula (0.15 points)
    # The original E25 had =SUM(E1:E24). It should still be a SUM.
    # This component only passes in golden because in the initial, E20 references E25
    # creating circular ref. Here we check E25 has SUM AND E20 doesn't reference E25
    # (compound check anchored to the change).
    try:
        if e25_val is not None and isinstance(e25_val, str) and e25_val.startswith('='):
            formula_upper = e25_val.upper()
            if 'SUM' in formula_upper:
                # Compound: E25 has SUM AND the circular ref is broken (E20 doesn't ref E25)
                e20_is_formula = (e20_val is not None and isinstance(e20_val, str) and
                                  e20_val.startswith('='))
                e20_no_e25_ref = e20_is_formula and 'E25' not in e20_val.upper()
                if e20_no_e25_ref:
                    print(f"PASS: Component 4 - E25 has SUM and circular ref broken (0.15 pts)")
                    total_score += 0.15
                else:
                    print(f"FAIL: Component 4 - E25 has SUM but circular ref still exists")
            else:
                print(f"FAIL: Component 4 - E25 does not contain SUM: {e25_val}")
        else:
            print(f"FAIL: Component 4 - E25 is not a formula: {repr(e25_val)}")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
