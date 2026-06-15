"""
Initial Setup: Budget spreadsheet with circular reference
Task ID: calc_tbl_013
Domain: libreoffice_calc

Creates a budget spreadsheet where E25=SUM(E1:E24) and E20 references E25,
creating a circular dependency.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_013'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget"

    # Headers in row 1 — but note E1:E24 are the budget amounts, so let's
    # use a separate header row approach. Actually the task says E25=SUM(E1:E24)
    # meaning E1 through E24 contain data. Let's put headers in A-D and have
    # column E be "Amount" with values from row 1 to row 24.
    # We'll use row 1 as headers and adjust: E1 is header, data E2:E24,
    # but task says E25=SUM(E1:E24). So E1 must have data.
    # Let's make no header row — or put headers in a different way.
    #
    # Simplest: A1:E1 are headers, data starts row 2. But task says E25=SUM(E1:E24).
    # So E1 must be a number. Let's have NO header row — just data rows 1-24, total row 25.
    # We can put column headers as column A=line item number, B=Category, C=Department,
    # D=Description, E=Amount.
    # Actually, let's just follow the task literally: rows 1-24 have budget data in column E,
    # row 25 has the total. Let's add headers in a freeze-pane friendly way by NOT using
    # headers (the task specifically references E1:E24 as data).

    # Column headers as a comment approach - let's use columns A-E with:
    # A = Line #, B = Category, C = Department, D = Description, E = Amount
    # But we need E1 to be data per the task spec. So no header row.
    # Let's just build 24 rows of budget data.

    # Budget line items
    budget_items = [
        ("Office Supplies", "Operations", "Printer paper, toner, pens", 2450),
        ("Software Licenses", "IT", "Annual renewal for dev tools", 18500),
        ("Travel - Q1", "Sales", "Client visits Jan-Mar", 8200),
        ("Training Programs", "HR", "Leadership development series", 5600),
        ("Cloud Infrastructure", "IT", "AWS hosting and services", 32000),
        ("Marketing Campaign", "Marketing", "Spring product launch", 15750),
        ("Equipment Maintenance", "Facilities", "HVAC and elevator service", 4300),
        ("Consulting Fees", "Finance", "External audit preparation", 12000),
        ("Employee Benefits", "HR", "Health plan adjustments", 28500),
        ("Research Materials", "R&D", "Lab supplies and journals", 6800),
        ("Client Entertainment", "Sales", "Dinners and event tickets", 3900),
        ("Security Upgrades", "IT", "Firewall and endpoint protection", 9200),
        ("Shipping & Logistics", "Operations", "Courier and freight costs", 7100),
        ("Legal Retainer", "Legal", "Outside counsel monthly fee", 11000),
        ("Utilities", "Facilities", "Electric, water, internet", 5400),
        ("Recruitment Costs", "HR", "Job postings and recruiter fees", 8900),
        ("Product Testing", "R&D", "QA lab equipment rental", 4200),
        ("Insurance Premiums", "Finance", "Liability and property coverage", 16300),
        ("Conference Sponsorship", "Marketing", "Annual industry summit", 7500),
        # Row 20 — this will have the circular reference formula
        (None, None, None, None),  # placeholder, will set formula below
        ("Vehicle Fleet", "Operations", "Fuel and maintenance", 6100),
        ("Charitable Donations", "Finance", "Community outreach program", 3000),
        ("Contingency Fund", "Finance", "Unplanned expense reserve", 10000),
        ("Year-End Bonus Pool", "HR", "Performance-based bonuses", 22000),
    ]

    # Write data rows 1-24
    for r, (cat, dept, desc, amt) in enumerate(budget_items, 1):
        if r == 20:
            # Row 20: "Remaining Budget" with circular reference
            ws.cell(row=r, column=1, value="Remaining Budget")
            ws.cell(row=r, column=2, value="Finance")
            ws.cell(row=r, column=3, value="Calculated from total minus allocated")
            # E20 references E25 — creating circular dependency
            ws.cell(row=r, column=5, value="=E25-SUM(E1:E19)")
        else:
            ws.cell(row=r, column=1, value=cat)
            ws.cell(row=r, column=2, value=dept)
            ws.cell(row=r, column=3, value=desc)
            ws.cell(row=r, column=5, value=amt)

    # Row 25: Total with SUM(E1:E24) — circular because E20 refs E25
    ws.cell(row=25, column=1, value="TOTAL")
    ws.cell(row=25, column=2, value="")
    ws.cell(row=25, column=3, value="Grand Total - All Departments")
    ws.cell(row=25, column=5, value="=SUM(E1:E24)")

    # Add a "D" column header-like label in D column (optional description)
    # Actually let's add column E header context via column D
    # Make row 25 bold for emphasis
    bold_font = Font(bold=True)
    for col in range(1, 6):
        ws.cell(row=25, column=col).font = bold_font

    # Column widths for readability
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 5
    ws.column_dimensions["E"].width = 15

    # Number format for amounts
    for r in range(1, 26):
        ws.cell(row=r, column=5).number_format = '#,##0'

    # Add a second sheet with department summary for complexity
    ws2 = wb.create_sheet("Departments")
    depts = [
        ("Operations", 3),
        ("IT", 3),
        ("Sales", 2),
        ("HR", 3),
        ("Marketing", 2),
        ("Facilities", 2),
        ("Finance", 3),
        ("R&D", 2),
        ("Legal", 1),
    ]
    ws2.cell(row=1, column=1, value="Department")
    ws2.cell(row=1, column=2, value="Line Items")
    ws2.cell(row=1, column=3, value="Notes")
    for i, (dept, count) in enumerate(depts, 2):
        ws2.cell(row=i, column=1, value=dept)
        ws2.cell(row=i, column=2, value=count)
        ws2.cell(row=i, column=3, value=f"Budget allocation for {dept}")

    # Bold header row on Departments sheet
    for col in range(1, 4):
        ws2.cell(row=1, column=col).font = Font(bold=True)

    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 30

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
