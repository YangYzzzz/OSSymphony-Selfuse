"""
Initial Setup: Build a project management index sheet navigation hub
Task ID: calc_gen_hyperlinks_056
Domain: libreoffice_calc

Creates a workbook with:
  - 'Index' sheet: project list with headers and data in columns A-E,
    columns F (Sheet Link) and G (Jira Link) intentionally EMPTY
  - 'ProjectA' through 'ProjectE' sheets with project detail data
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_hyperlinks_056'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # ─── Index Sheet ────────────────────────────────────────────────────────────
    ws_index = wb.active
    ws_index.title = 'Index'

    # Headers in row 1
    headers = ['Project', 'Ticket ID', 'Status', 'Owner', 'Due Date', 'Sheet Link', 'Jira Link']
    header_font = Font(name='Calibri', size=12, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, 1):
        cell = ws_index.cell(row=1, column=col_idx, value=h)
        cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFFFF')
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # Project data rows 2-6 — columns A-E only, F and G intentionally empty
    projects = [
        ('ProjectA', 'PROJ-101', 'Complete',    'Sarah Chen',     '2025-03-15'),
        ('ProjectB', 'PROJ-242', 'In Progress', 'Marcus Johnson', '2025-04-30'),
        ('ProjectC', 'PROJ-387', 'Blocked',     'Elena Rodriguez','2025-05-20'),
        ('ProjectD', 'PROJ-512', 'In Progress', 'David Kim',      '2025-06-10'),
        ('ProjectE', 'PROJ-634', 'Complete',    'Priya Nair',     '2025-07-01'),
    ]

    row_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row_idx, (proj, ticket, status, owner, due) in enumerate(projects, 2):
        ws_index.cell(row=row_idx, column=1, value=proj).border = row_border
        ws_index.cell(row=row_idx, column=2, value=ticket).border = row_border
        ws_index.cell(row=row_idx, column=3, value=status).border = row_border
        ws_index.cell(row=row_idx, column=4, value=owner).border = row_border
        ws_index.cell(row=row_idx, column=5, value=due).border = row_border
        # Columns F and G: borders only, NO value and NO formula
        ws_index.cell(row=row_idx, column=6).border = row_border
        ws_index.cell(row=row_idx, column=7).border = row_border

    # Column widths for readability
    ws_index.column_dimensions['A'].width = 14
    ws_index.column_dimensions['B'].width = 12
    ws_index.column_dimensions['C'].width = 14
    ws_index.column_dimensions['D'].width = 18
    ws_index.column_dimensions['E'].width = 12
    ws_index.column_dimensions['F'].width = 14
    ws_index.column_dimensions['G'].width = 30
    ws_index.row_dimensions[1].height = 22

    # ─── Project Detail Sheets ──────────────────────────────────────────────────
    project_details = {
        'ProjectA': {
            'title': 'Alpha Platform Migration',
            'description': 'Migrate legacy monolith to cloud-native microservices architecture.',
            'tasks': [
                ('T-001', 'Architecture design', 'Done',        'Sarah Chen',     '2025-01-20', 'High'),
                ('T-002', 'Data schema mapping', 'Done',        'Marcus Johnson', '2025-01-28', 'High'),
                ('T-003', 'API gateway setup',   'Done',        'Elena Rodriguez','2025-02-10', 'Medium'),
                ('T-004', 'Service deployment',  'Done',        'David Kim',      '2025-02-25', 'High'),
                ('T-005', 'Integration testing', 'Done',        'Priya Nair',     '2025-03-05', 'High'),
                ('T-006', 'Load testing',        'Done',        'Sarah Chen',     '2025-03-10', 'Medium'),
                ('T-007', 'Documentation',       'Done',        'Marcus Johnson', '2025-03-14', 'Low'),
                ('T-008', 'Stakeholder sign-off','Done',        'Sarah Chen',     '2025-03-15', 'Medium'),
            ],
        },
        'ProjectB': {
            'title': 'Beta Analytics Dashboard',
            'description': 'Build real-time analytics dashboard for executive reporting.',
            'tasks': [
                ('T-001', 'Requirements gathering', 'Done',        'Marcus Johnson', '2025-02-01', 'High'),
                ('T-002', 'Data pipeline design',   'Done',        'Elena Rodriguez','2025-02-15', 'High'),
                ('T-003', 'Backend API',             'In Progress', 'David Kim',      '2025-03-20', 'High'),
                ('T-004', 'Dashboard UI',            'In Progress', 'Priya Nair',     '2025-04-01', 'High'),
                ('T-005', 'User auth integration',   'Not Started', 'Sarah Chen',     '2025-04-15', 'Medium'),
                ('T-006', 'UAT testing',             'Not Started', 'Marcus Johnson', '2025-04-25', 'Medium'),
                ('T-007', 'Performance tuning',      'Not Started', 'Elena Rodriguez','2025-04-28', 'Medium'),
                ('T-008', 'Production rollout',      'Not Started', 'Marcus Johnson', '2025-04-30', 'High'),
            ],
        },
        'ProjectC': {
            'title': 'Compliance Audit System',
            'description': 'Automated compliance tracking and audit trail generation.',
            'tasks': [
                ('T-001', 'Regulatory analysis',  'Done',    'Elena Rodriguez','2025-01-10', 'High'),
                ('T-002', 'System design',        'Done',    'David Kim',      '2025-01-25', 'High'),
                ('T-003', 'Database schema',      'Done',    'Priya Nair',     '2025-02-05', 'High'),
                ('T-004', 'Audit trail module',   'Blocked', 'Sarah Chen',     '2025-03-01', 'High'),
                ('T-005', 'Reporting engine',     'Blocked', 'Marcus Johnson', '2025-03-15', 'High'),
                ('T-006', 'Security review',      'Blocked', 'Elena Rodriguez','2025-04-01', 'High'),
                ('T-007', 'User training',        'Blocked', 'David Kim',      '2025-05-10', 'Medium'),
                ('T-008', 'Go-live approval',     'Blocked', 'Elena Rodriguez','2025-05-20', 'High'),
            ],
        },
        'ProjectD': {
            'title': 'DevOps Pipeline Overhaul',
            'description': 'Modernize CI/CD pipelines for all product teams.',
            'tasks': [
                ('T-001', 'Current state audit',  'Done',        'David Kim',      '2025-02-05', 'Medium'),
                ('T-002', 'Toolchain evaluation', 'Done',        'Priya Nair',     '2025-02-20', 'Medium'),
                ('T-003', 'Pipeline templates',   'In Progress', 'Sarah Chen',     '2025-03-10', 'High'),
                ('T-004', 'Team onboarding',      'In Progress', 'Marcus Johnson', '2025-04-01', 'High'),
                ('T-005', 'Security hardening',   'Not Started', 'Elena Rodriguez','2025-05-01', 'High'),
                ('T-006', 'Monitoring setup',     'Not Started', 'David Kim',      '2025-05-20', 'Medium'),
                ('T-007', 'Runbook creation',     'Not Started', 'Priya Nair',     '2025-06-05', 'Low'),
                ('T-008', 'Final sign-off',       'Not Started', 'David Kim',      '2025-06-10', 'Medium'),
            ],
        },
        'ProjectE': {
            'title': 'Employee Portal Redesign',
            'description': 'Redesign self-service HR portal with modern UX.',
            'tasks': [
                ('T-001', 'UX research',          'Done', 'Priya Nair',     '2025-01-15', 'Medium'),
                ('T-002', 'Wireframe design',     'Done', 'Sarah Chen',     '2025-02-01', 'Medium'),
                ('T-003', 'UI component library', 'Done', 'Marcus Johnson', '2025-03-01', 'High'),
                ('T-004', 'Backend integration',  'Done', 'Elena Rodriguez','2025-04-15', 'High'),
                ('T-005', 'Accessibility audit',  'Done', 'David Kim',      '2025-05-01', 'Medium'),
                ('T-006', 'QA testing',           'Done', 'Priya Nair',     '2025-06-01', 'High'),
                ('T-007', 'Pilot rollout',        'Done', 'Sarah Chen',     '2025-06-20', 'High'),
                ('T-008', 'Full deployment',      'Done', 'Priya Nair',     '2025-07-01', 'High'),
            ],
        },
    }

    for sheet_name, detail in project_details.items():
        ws = wb.create_sheet(sheet_name)

        # Sheet title
        ws['A1'] = detail['title']
        ws['A1'].font = Font(name='Calibri', size=14, bold=True)
        ws['A2'] = detail['description']
        ws['A2'].font = Font(name='Calibri', size=10, italic=True)

        # Task table headers — row 4
        task_headers = ['Task ID', 'Description', 'Status', 'Assignee', 'Due Date', 'Priority']
        th_font = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
        th_fill = PatternFill(start_color='FF375623', end_color='FF375623', fill_type='solid')
        for c_idx, h in enumerate(task_headers, 1):
            cell = ws.cell(row=4, column=c_idx, value=h)
            cell.font = th_font
            cell.fill = th_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = row_border

        # Task rows starting at row 5
        for r_off, task in enumerate(detail['tasks'], 5):
            for c_idx, val in enumerate(task, 1):
                cell = ws.cell(row=r_off, column=c_idx, value=val)
                cell.border = row_border

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 10
        ws.row_dimensions[1].height = 22

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets:', wb.sheetnames)
    print('Index columns F and G are intentionally empty (no formulas/hyperlinks)')


create_initial()
