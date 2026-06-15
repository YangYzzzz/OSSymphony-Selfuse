"""
Reward Script: Build a project management index sheet with hyperlinks and conditional formatting
Task ID: calc_gen_hyperlinks_056
Domain: libreoffice_calc
Scoring:
  - Component 1: Sheet-link HYPERLINK formulas in F2:F6 (0.4 points)
  - Component 2: Jira-link HYPERLINK formulas in G2:G6 (0.4 points)
  - Component 3: Conditional formatting on Status column C2:C6 (0.2 points)
Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_hyperlinks_056'

# Expected project names and their corresponding sheet names
PROJECTS = ['ProjectA', 'ProjectB', 'ProjectC', 'ProjectD', 'ProjectE']
JIRA_BASE = 'https://company.atlassian.net/browse/'


def normalize_formula(formula):
    """Normalize formula for comparison: strip whitespace, uppercase function names."""
    if not isinstance(formula, str):
        return ''
    return formula.strip()


def check_sheet_hyperlink_formula(formula, project_name):
    """
    Check if a formula is a valid HYPERLINK to the project sheet.
    Valid patterns: =HYPERLINK("#'ProjectA'.A1","Open Sheet") or variants with '#ProjectA.A1'
    """
    if not formula:
        return False
    f = formula.upper().replace(' ', '')
    # Must start with =HYPERLINK
    if not f.startswith('=HYPERLINK('):
        return False
    # Must contain a reference to the project name (case-insensitive)
    if project_name.upper() not in f:
        return False
    # Must start with '#' anchor for internal sheet link
    if '"#' not in formula and "'#" not in formula:
        # Check if # is part of the formula URL
        if '#' not in formula:
            return False
    return True


def check_jira_hyperlink_formula(formula, row_index):
    """
    Check if a formula is a valid HYPERLINK constructing Jira URL.
    Expected pattern: =HYPERLINK("https://company.atlassian.net/browse/"&B2,B2)
    row_index is the actual row number (2-6)
    """
    if not formula:
        return False
    f = formula.upper().replace(' ', '')
    # Must start with =HYPERLINK
    if not f.startswith('=HYPERLINK('):
        return False
    # Must contain the jira base URL
    jira_upper = JIRA_BASE.upper()
    if jira_upper not in formula.upper():
        return False
    # Must reference the B column cell for this row (concatenation)
    b_ref = f'B{row_index}'
    if b_ref not in formula.upper():
        return False
    # Must use & for concatenation
    if '&' not in formula:
        return False
    return True


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify Index sheet exists
    if 'Index' not in wb.sheetnames:
        print("CRITICAL: 'Index' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Index']

    # -----------------------------------------------------------------------
    # Component 1: Sheet-link HYPERLINK formulas in F2:F6 (0.4 points)
    # Each row should have =HYPERLINK("#'ProjectX'.A1","Open Sheet")
    # Award 0.08 per correct row (5 rows × 0.08 = 0.40)
    # -----------------------------------------------------------------------
    comp1_points = 0.0
    comp1_per_row = 0.4 / 5
    print("--- Component 1: Sheet-link HYPERLINK formulas (F2:F6) ---")
    for i, project in enumerate(PROJECTS):
        row = i + 2  # rows 2-6
        cell = ws.cell(row=row, column=6)
        formula = cell.value
        if check_sheet_hyperlink_formula(formula, project):
            print(f"  PASS: F{row} has valid sheet hyperlink formula: {repr(formula)}")
            comp1_points += comp1_per_row
        else:
            print(f"  FAIL: F{row} expected HYPERLINK to {project}, found: {repr(formula)}")

    total_score += comp1_points
    print(f"Component 1 score: {comp1_points:.2f}/0.40\n")

    # -----------------------------------------------------------------------
    # Component 2: Jira-link HYPERLINK formulas in G2:G6 (0.4 points)
    # Each row should have =HYPERLINK("https://company.atlassian.net/browse/"&B2,B2)
    # Award 0.08 per correct row (5 rows × 0.08 = 0.40)
    # -----------------------------------------------------------------------
    comp2_points = 0.0
    comp2_per_row = 0.4 / 5
    print("--- Component 2: Jira-link HYPERLINK formulas (G2:G6) ---")
    for i, project in enumerate(PROJECTS):
        row = i + 2  # rows 2-6
        cell = ws.cell(row=row, column=7)
        formula = cell.value
        if check_jira_hyperlink_formula(formula, row):
            print(f"  PASS: G{row} has valid Jira hyperlink formula: {repr(formula)}")
            comp2_points += comp2_per_row
        else:
            print(f"  FAIL: G{row} expected Jira HYPERLINK with B{row} reference, found: {repr(formula)}")

    total_score += comp2_points
    print(f"Component 2 score: {comp2_points:.2f}/0.40\n")

    # -----------------------------------------------------------------------
    # Component 3: Conditional formatting on Status column C2:C6 (0.2 points)
    # Must have rules for: Complete (green), In Progress (yellow), Blocked (red)
    # Award 0.067 per matching rule (3 rules × 0.067 ≈ 0.20)
    # -----------------------------------------------------------------------
    comp3_points = 0.0
    comp3_per_rule = 0.2 / 3
    print("--- Component 3: Conditional formatting on Status column (C2:C6) ---")

    # Check if any conditional formatting exists
    cf_rules_found = list(ws.conditional_formatting)
    if not cf_rules_found:
        print("  FAIL: No conditional formatting found on Index sheet")
    else:
        # Collect all rules across all CF ranges that overlap C2:C6
        all_rules = []
        for cf_range in cf_rules_found:
            cf_str = str(cf_range)
            # Check if this CF range covers C column rows 2-6
            if 'C' in cf_str:
                for rule in ws.conditional_formatting[cf_range]:
                    all_rules.append(rule)

        if not all_rules:
            print("  FAIL: No conditional formatting found for column C")
        else:
            print(f"  Found {len(all_rules)} conditional formatting rule(s) for column C")

            # Check for Complete -> green fill
            complete_found = False
            in_progress_found = False
            blocked_found = False

            for rule in all_rules:
                formulas = getattr(rule, 'formula', [])
                rule_formula_str = ' '.join(str(f) for f in formulas) if formulas else ''
                rule_type = getattr(rule, 'type', '')
                rule_op = getattr(rule, 'operator', '')

                # Try to get fill color
                fill_color = None
                try:
                    if rule.dxf and rule.dxf.fill:
                        fill_color = rule.dxf.fill.fgColor.rgb
                except Exception:
                    pass

                print(f"  Rule: type={rule_type}, op={rule_op}, formula={rule_formula_str}, fill={fill_color}")

                # Complete -> green (check fill color is greenish or formula contains 'Complete')
                if 'Complete' in rule_formula_str or '"complete"' in rule_formula_str.lower():
                    if fill_color:
                        # Accept any greenish color (green-dominant channel)
                        try:
                            argb = fill_color.lstrip('0') if len(fill_color) > 6 else fill_color
                            # FF00B050 = green
                            r_val = int(fill_color[2:4], 16) if len(fill_color) == 8 else 255
                            g_val = int(fill_color[4:6], 16) if len(fill_color) == 8 else 0
                            b_val = int(fill_color[6:8], 16) if len(fill_color) == 8 else 0
                            if g_val > r_val and g_val > b_val:
                                print(f"    PASS: Complete rule with green fill ({fill_color})")
                                complete_found = True
                        except Exception as e:
                            print(f"    WARN: Could not parse fill color {fill_color}: {e}")
                    else:
                        # No fill color but has the formula — partial credit
                        print(f"    PARTIAL: Complete rule found but fill color unclear")
                        complete_found = True

                # In Progress -> yellow (yellowish fill)
                if 'In Progress' in rule_formula_str or '"in progress"' in rule_formula_str.lower():
                    if fill_color:
                        try:
                            r_val = int(fill_color[2:4], 16) if len(fill_color) == 8 else 255
                            g_val = int(fill_color[4:6], 16) if len(fill_color) == 8 else 0
                            b_val = int(fill_color[6:8], 16) if len(fill_color) == 8 else 0
                            # Yellow: high R and G, B lower than both R and G
                            if r_val > 150 and g_val > 150 and b_val < r_val and b_val < g_val:
                                print(f"    PASS: In Progress rule with yellow fill ({fill_color})")
                                in_progress_found = True
                        except Exception as e:
                            print(f"    WARN: Could not parse fill color {fill_color}: {e}")
                    else:
                        print(f"    PARTIAL: In Progress rule found but fill color unclear")
                        in_progress_found = True

                # Blocked -> red (reddish fill)
                if 'Blocked' in rule_formula_str or '"blocked"' in rule_formula_str.lower():
                    if fill_color:
                        try:
                            r_val = int(fill_color[2:4], 16) if len(fill_color) == 8 else 255
                            g_val = int(fill_color[4:6], 16) if len(fill_color) == 8 else 0
                            b_val = int(fill_color[6:8], 16) if len(fill_color) == 8 else 0
                            # Red: high R, lower G and B
                            if r_val > 150 and r_val > g_val and r_val > b_val:
                                print(f"    PASS: Blocked rule with red fill ({fill_color})")
                                blocked_found = True
                        except Exception as e:
                            print(f"    WARN: Could not parse fill color {fill_color}: {e}")
                    else:
                        print(f"    PARTIAL: Blocked rule found but fill color unclear")
                        blocked_found = True

            if complete_found:
                comp3_points += comp3_per_rule
            else:
                print("  FAIL: No green fill rule for 'Complete' status found")

            if in_progress_found:
                comp3_points += comp3_per_rule
            else:
                print("  FAIL: No yellow fill rule for 'In Progress' status found")

            if blocked_found:
                comp3_points += comp3_per_rule
            else:
                print("  FAIL: No red fill rule for 'Blocked' status found")

    total_score += comp3_points
    print(f"Component 3 score: {comp3_points:.3f}/0.200\n")

    final_score = round(min(total_score, 1.0), 4)
    print(f"Score: {total_score:.4f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
