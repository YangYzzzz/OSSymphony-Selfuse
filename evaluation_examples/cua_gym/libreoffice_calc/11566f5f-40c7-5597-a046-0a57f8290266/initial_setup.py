"""
Initial Setup: Product comparison spreadsheet with no borders
Task ID: calc_gsd_003
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_003'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Comparison'

    # Headers
    headers = ['Product', 'Price', 'Rating', 'Stock', 'Category']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 9 rows of realistic product data
    data = [
        ['Wireless Earbuds Pro',       79.99, 4.5, 230, 'Electronics'],
        ['Organic Green Tea',          12.49, 4.8, 540, 'Beverages'],
        ['Stainless Steel Water Bottle', 24.95, 4.3, 185, 'Kitchen'],
        ['Bamboo Cutting Board',       18.50, 4.6, 320, 'Kitchen'],
        ['LED Desk Lamp',              45.00, 4.2, 112, 'Electronics'],
        ['Yoga Mat Premium',           34.99, 4.7, 275, 'Fitness'],
        ['Ceramic Coffee Mug Set',     29.99, 4.4, 198, 'Kitchen'],
        ['Bluetooth Speaker Mini',     55.00, 4.1, 89,  'Electronics'],
        ['Natural Lip Balm Pack',       9.99, 4.9, 610, 'Beauty'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths for readability
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 14

    # No borders applied - that is the task
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
