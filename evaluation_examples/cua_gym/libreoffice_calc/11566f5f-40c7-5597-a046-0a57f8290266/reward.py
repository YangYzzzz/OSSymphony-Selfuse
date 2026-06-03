"""
Reward Script: Apply borders to product comparison table
Task ID: calc_gsd_003
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.35): Data cells A2:E10 have thin borders on all sides
  - Component 2 (0.35): Header row A1:E1 has thick top and bottom borders
  - Component 3 (0.20): Header row outer left (A1) and right (E1) are thick
  - Component 4 (0.10): Interior header cell left/right borders are thin (not thick)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_003'


def persist_app_state(domain: str):
    """Attempt to save any unsaved LibreOffice edits via Ctrl+S."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet 'Comparison' must exist
    if 'Comparison' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Comparison' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Comparison']

    # ---------------------------------------------------------------
    # Component 1: Data cells A2:E10 have thin borders on all 4 sides (0.35 pts)
    # This checks that every data cell got an outside/box border applied.
    # In initial_env, borders are None everywhere, so this FAILS on initial.
    # ---------------------------------------------------------------
    try:
        data_cells_total = 0
        data_cells_pass = 0
        for row in range(2, 11):  # rows 2-10
            for col in range(1, 6):  # columns A-E
                data_cells_total += 1
                cell = ws.cell(row=row, column=col)
                b = cell.border
                # All 4 sides must have a non-None border style
                if (b.left.style is not None and
                    b.right.style is not None and
                    b.top.style is not None and
                    b.bottom.style is not None):
                    data_cells_pass += 1

        if data_cells_total > 0:
            ratio = data_cells_pass / data_cells_total
            if ratio == 1.0:
                print(f"PASS: Component 1 — All {data_cells_total} data cells have borders on all sides (0.35 pts)")
                total_score += 0.35
            elif ratio >= 0.8:
                partial = round(0.35 * ratio, 2)
                print(f"PARTIAL: Component 1 — {data_cells_pass}/{data_cells_total} data cells have full borders ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 1 — Only {data_cells_pass}/{data_cells_total} data cells have borders on all sides")
        else:
            print("FAIL: Component 1 — No data cells found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ---------------------------------------------------------------
    # Component 2: Header row A1:E1 has thick top AND thick bottom borders (0.35 pts)
    # The task requires a thick outer border around the header row.
    # In initial_env, all borders are None, so this FAILS on initial.
    # ---------------------------------------------------------------
    try:
        header_thick_count = 0
        header_cells_total = 5  # A1 through E1
        for col in range(1, 6):
            cell = ws.cell(row=1, column=col)
            b = cell.border
            # Top and bottom of header cells must be thick (or medium — accept both)
            top_ok = b.top.style in ('thick', 'medium')
            bot_ok = b.bottom.style in ('thick', 'medium')
            if top_ok and bot_ok:
                header_thick_count += 1

        if header_thick_count == header_cells_total:
            print(f"PASS: Component 2 — All 5 header cells have thick top+bottom borders (0.35 pts)")
            total_score += 0.35
        elif header_thick_count >= 3:
            partial = round(0.35 * (header_thick_count / header_cells_total), 2)
            print(f"PARTIAL: Component 2 — {header_thick_count}/{header_cells_total} header cells have thick top+bottom ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {header_thick_count}/{header_cells_total} header cells have thick top+bottom borders")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---------------------------------------------------------------
    # Component 3: Header outer edges — A1 left is thick, E1 right is thick (0.20 pts)
    # This verifies the thick BOX border wraps around the header row entirely.
    # In initial_env, borders are None, so this FAILS on initial.
    # ---------------------------------------------------------------
    try:
        a1_border = ws['A1'].border
        e1_border = ws['E1'].border
        a1_left_ok = a1_border.left.style in ('thick', 'medium')
        e1_right_ok = e1_border.right.style in ('thick', 'medium')

        if a1_left_ok and e1_right_ok:
            print(f"PASS: Component 3 — A1 left={a1_border.left.style}, E1 right={e1_border.right.style} (0.20 pts)")
            total_score += 0.20
        elif a1_left_ok or e1_right_ok:
            print(f"PARTIAL: Component 3 — A1 left={a1_border.left.style}, E1 right={e1_border.right.style} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — A1 left={a1_border.left.style}, E1 right={e1_border.right.style}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ---------------------------------------------------------------
    # Component 4: Interior header borders (B1-D1) left+right are thin, not thick (0.10 pts)
    # This confirms that only the OUTER boundary of the header row is thick,
    # while internal cell dividers remain thin.
    # In initial_env, borders are None, so this FAILS on initial (None != thin).
    # ---------------------------------------------------------------
    try:
        interior_ok = 0
        interior_total = 3  # B1, C1, D1
        for col_letter in ['B', 'C', 'D']:
            cell = ws[f'{col_letter}1']
            b = cell.border
            # Left and right should be thin (not thick, not None)
            if b.left.style == 'thin' and b.right.style == 'thin':
                interior_ok += 1

        if interior_ok == interior_total:
            print(f"PASS: Component 4 — Interior header cells B1:D1 have thin left/right borders (0.10 pts)")
            total_score += 0.10
        elif interior_ok > 0:
            partial = round(0.10 * (interior_ok / interior_total), 2)
            print(f"PARTIAL: Component 4 — {interior_ok}/{interior_total} interior header cells correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Interior header cells do not have thin left/right borders")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
