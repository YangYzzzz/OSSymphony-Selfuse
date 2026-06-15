"""
Initial Setup: Regional sales spreadsheet with 19 products and 5 regions.
Task ID: calc_gcv_041
Domain: libreoffice_calc
No conditional formatting applied - that is the agent's task.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_041'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Regional_Sales"

    # --- Headers ---
    headers = ["Product Name", "North", "South", "East", "West", "Central"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # --- Product data: 19 products x 5 regions ---
    products = [
        ["Wireless Earbuds Pro",       4520, 3890, 5120, 2780, 4100],
        ["Smart Home Hub",             3150, 4230, 2870, 3960, 3540],
        ["Portable Charger 20K",       6780, 5430, 6780, 7120, 5890],
        ["Bluetooth Speaker Mini",     2340, 3670, 2890, 2340, 4520],
        ["Fitness Tracker Band",       5890, 5890, 4320, 5670, 3210],
        ["USB-C Docking Station",      1890, 2450, 3780, 2100, 2670],
        ["Noise Canceling Headphones", 8920, 7650, 9340, 8150, 7890],
        ["Mechanical Keyboard RGB",    3450, 4120, 3780, 5230, 4560],
        ["Webcam 4K Ultra",            2670, 1980, 2340, 3120, 2890],
        ["External SSD 1TB",           7230, 6890, 7450, 6230, 8120],
        ["Smart Watch Elite",          9120, 8760, 7890, 9450, 8340],
        ["Wireless Mouse Ergonomic",   1560, 2340, 1890, 1780, 2120],
        ["Monitor Light Bar",          4230, 3560, 4890, 3890, 4230],
        ["Tablet Stand Adjustable",    1230, 1670, 1450, 980,  1890],
        ["Power Strip Smart",          3670, 4560, 3120, 4230, 3890],
        ["Cable Management Kit",       890,  1230, 1560, 1120, 780],
        ["Desk Organizer Premium",     2560, 2890, 2120, 3450, 2780],
        ["Screen Protector Pack",      670,  890,  1230, 780,  560],
        ["Laptop Cooling Pad",         3890, 4230, 3560, 4670, 3120],
    ]

    data_align = Alignment(horizontal="center", vertical="center")
    name_align = Alignment(horizontal="left", vertical="center")
    num_format = '#,##0'

    for r, row_data in enumerate(products, 2):
        # Product name (col A)
        cell = ws.cell(row=r, column=1, value=row_data[0])
        cell.alignment = name_align
        cell.border = border

        # Sales figures (cols B-F)
        for c, val in enumerate(row_data[1:], 2):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = data_align
            cell.border = border
            cell.number_format = num_format

    # --- Column widths ---
    ws.column_dimensions["A"].width = 28
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 14

    # --- Freeze header row ---
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
