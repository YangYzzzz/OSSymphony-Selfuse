"""
Initial Setup: Move Q1 and Q2 sheets to chronological order
Task ID: calc_gsi_028
Domain: libreoffice_calc

Creates a workbook with sheets in order: Summary, Q3, Q4, Q1, Q2
Each sheet contains realistic quarterly financial data.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_028'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # Style definitions
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    currency_fmt = '$#,##0.00'
    pct_fmt = '0.0%'
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # --- Helper to populate a quarterly sheet ---
    def populate_quarter(ws, quarter_label, data_rows):
        headers = ['Department', 'Revenue', 'Expenses', 'Net Profit', 'Growth Rate', 'Headcount']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for r, row_data in enumerate(data_rows, 2):
            for c, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = thin_border
                if c in (2, 3, 4):
                    cell.number_format = currency_fmt
                elif c == 5:
                    cell.number_format = pct_fmt

        # Column widths
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 12

    # --- Sheet 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = 'Summary'

    summary_headers = ['Quarter', 'Total Revenue', 'Total Expenses', 'Total Net Profit', 'Avg Growth']
    for col, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    summary_data = [
        ['Q1 2025', 2845000, 1923500, 921500, 0.042],
        ['Q2 2025', 3012000, 2054000, 958000, 0.058],
        ['Q3 2025', 2967000, 2015600, 951400, 0.035],
        ['Q4 2025', 3254000, 2187300, 1066700, 0.072],
    ]
    for r, row_data in enumerate(summary_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_summary.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c in (2, 3, 4):
                cell.number_format = currency_fmt
            elif c == 5:
                cell.number_format = pct_fmt

    ws_summary.column_dimensions['A'].width = 14
    ws_summary.column_dimensions['B'].width = 18
    ws_summary.column_dimensions['C'].width = 18
    ws_summary.column_dimensions['D'].width = 20
    ws_summary.column_dimensions['E'].width = 14

    # --- Sheet 2: Q3 (intentionally out of order) ---
    ws_q3 = wb.create_sheet('Q3')
    q3_data = [
        ['Engineering', 845000, 562000, 283000, 0.031, 48],
        ['Sales', 723000, 498000, 225000, 0.028, 35],
        ['Marketing', 412000, 315000, 97000, 0.045, 22],
        ['Operations', 298000, 225600, 72400, 0.018, 18],
        ['Finance', 187000, 142000, 45000, 0.025, 12],
        ['HR', 156000, 118000, 38000, 0.033, 10],
        ['Legal', 134000, 98000, 36000, 0.022, 8],
        ['R&D', 112000, 87000, 25000, 0.062, 15],
        ['Customer Support', 67000, 52000, 15000, 0.019, 14],
        ['IT Infrastructure', 33000, 18000, 15000, 0.041, 9],
    ]
    populate_quarter(ws_q3, 'Q3', q3_data)

    # --- Sheet 3: Q4 ---
    ws_q4 = wb.create_sheet('Q4')
    q4_data = [
        ['Engineering', 928000, 608000, 320000, 0.068, 52],
        ['Sales', 812000, 543000, 269000, 0.075, 38],
        ['Marketing', 445000, 328000, 117000, 0.082, 24],
        ['Operations', 321000, 238300, 82700, 0.055, 19],
        ['Finance', 198000, 148000, 50000, 0.048, 12],
        ['HR', 167000, 124000, 43000, 0.061, 11],
        ['Legal', 142000, 102000, 40000, 0.038, 8],
        ['R&D', 125000, 94000, 31000, 0.078, 16],
        ['Customer Support', 72000, 55000, 17000, 0.065, 15],
        ['IT Infrastructure', 44000, 27000, 17000, 0.072, 10],
    ]
    populate_quarter(ws_q4, 'Q4', q4_data)

    # --- Sheet 4: Q1 ---
    ws_q1 = wb.create_sheet('Q1')
    q1_data = [
        ['Engineering', 812000, 548000, 264000, 0.038, 45],
        ['Sales', 689000, 472000, 217000, 0.042, 32],
        ['Marketing', 395000, 298000, 97000, 0.051, 20],
        ['Operations', 285000, 218500, 66500, 0.029, 17],
        ['Finance', 178000, 135000, 43000, 0.035, 11],
        ['HR', 148000, 112000, 36000, 0.038, 9],
        ['Legal', 128000, 94000, 34000, 0.027, 7],
        ['R&D', 105000, 82000, 23000, 0.055, 14],
        ['Customer Support', 62000, 48000, 14000, 0.024, 13],
        ['IT Infrastructure', 43000, 16000, 27000, 0.048, 8],
    ]
    populate_quarter(ws_q1, 'Q1', q1_data)

    # --- Sheet 5: Q2 ---
    ws_q2 = wb.create_sheet('Q2')
    q2_data = [
        ['Engineering', 862000, 578000, 284000, 0.055, 47],
        ['Sales', 745000, 510000, 235000, 0.063, 34],
        ['Marketing', 423000, 312000, 111000, 0.068, 21],
        ['Operations', 302000, 228000, 74000, 0.042, 18],
        ['Finance', 185000, 140000, 45000, 0.041, 12],
        ['HR', 155000, 116000, 39000, 0.045, 10],
        ['Legal', 132000, 97000, 35000, 0.032, 8],
        ['R&D', 110000, 85000, 25000, 0.065, 15],
        ['Customer Support', 65000, 50000, 15000, 0.038, 14],
        ['IT Infrastructure', 33000, 18000, 15000, 0.052, 9],
    ]
    populate_quarter(ws_q2, 'Q2', q2_data)

    # Verify sheet order is: Summary, Q3, Q4, Q1, Q2
    print(f'Sheet order: {wb.sheetnames}')
    assert wb.sheetnames == ['Summary', 'Q3', 'Q4', 'Q1', 'Q2'], \
        f'Expected [Summary, Q3, Q4, Q1, Q2], got {wb.sheetnames}'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
