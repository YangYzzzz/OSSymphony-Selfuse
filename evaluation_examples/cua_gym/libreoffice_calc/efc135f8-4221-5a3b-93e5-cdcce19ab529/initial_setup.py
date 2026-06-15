"""
Initial Setup: Comprehensive expense audit task
Task ID: osworld_multi_apps_doc_pdf_calc_008
Domain: libreoffice_calc (multi-app: calc + writer + pdf)

Creates:
  - /home/user/monthly_receipts/ directory with 6 PDF receipt files
  - /home/user/Desktop/expense_budget.ods with Category, Budget, Actual (empty),
    Variance (empty), Over_Budget (empty) columns
"""

import os
import shlex
import subprocess
import time

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_doc_pdf_calc_008'
BUDGET_FILE = f'{WORKDIR}/Desktop/expense_budget.ods'
RECEIPTS_DIR = f'{WORKDIR}/monthly_receipts'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_pdf_receipt(filepath, title, vendor, date, items, total, category):
    """Create a simple PDF receipt using fpdf2."""
    try:
        from fpdf import FPDF

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 12, title, ln=True, align="C")
        pdf.ln(4)

        pdf.set_font("Helvetica", "", 11)
        pdf.cell(0, 8, f"Vendor: {vendor}", ln=True)
        pdf.cell(0, 8, f"Date: {date}", ln=True)
        pdf.cell(0, 8, f"Category: {category}", ln=True)
        pdf.ln(4)

        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(110, 8, "Description", border=1)
        pdf.cell(40, 8, "Qty", border=1, align="C")
        pdf.cell(40, 8, "Amount", border=1, align="R", ln=True)

        pdf.set_font("Helvetica", "", 11)
        for desc, qty, amount in items:
            pdf.cell(110, 8, desc, border=1)
            pdf.cell(40, 8, str(qty), border=1, align="C")
            pdf.cell(40, 8, f"${amount:.2f}", border=1, align="R", ln=True)

        pdf.ln(4)
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(150, 10, "TOTAL:", align="R")
        pdf.cell(40, 10, f"${total:.2f}", align="R", ln=True)

        pdf.ln(6)
        pdf.set_font("Helvetica", "I", 9)
        pdf.cell(0, 8, "Thank you for your business. Please retain this receipt for your records.", ln=True, align="C")

        pdf.output(filepath)
        print(f"  Created PDF: {filepath}")
    except Exception as e:
        # Fallback: create a minimal text-based PDF manually
        print(f"  fpdf2 error: {e}. Creating minimal PDF.")
        # Minimal valid PDF
        content_stream = (
            f"BT\n"
            f"/F1 14 Tf\n"
            f"50 780 Td\n"
            f"({title}) Tj\n"
            f"0 -20 Td\n"
            f"(Vendor: {vendor}) Tj\n"
            f"0 -16 Td\n"
            f"(Date: {date}) Tj\n"
            f"0 -16 Td\n"
            f"(Category: {category}) Tj\n"
            f"0 -16 Td\n"
            f"(Total: ${total:.2f}) Tj\n"
            f"ET\n"
        )
        content_bytes = content_stream.encode('latin-1')
        pdf_content = (
            b"%PDF-1.4\n"
            b"1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n"
            b"2 0 obj\n<< /Type /Pages /Kids [3 0 R] /Count 1 >>\nendobj\n"
            b"3 0 obj\n<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792]\n"
            b"   /Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >>\nendobj\n"
            + b"4 0 obj\n<< /Length " + str(len(content_bytes)).encode() + b" >>\nstream\n"
            + content_bytes + b"\nendstream\nendobj\n"
            b"5 0 obj\n<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>\nendobj\n"
            b"xref\n0 6\n0000000000 65535 f \n"
        )
        with open(filepath, 'wb') as f:
            f.write(pdf_content)


def create_receipts():
    """Create the monthly_receipts directory with 6 PDF files."""
    os.makedirs(RECEIPTS_DIR, exist_ok=True)

    receipts = [
        {
            "filename": "receipt_travel_airfare.pdf",
            "title": "BUSINESS TRAVEL RECEIPT",
            "vendor": "SkyWings Airlines",
            "date": "2025-03-05",
            "category": "Travel",
            "items": [
                ("Round-trip flight SFO-NYC (Mar 10-14)", 1, 890.00),
                ("Checked baggage fee (2 bags)", 2, 35.00),
                ("Airport transfer - taxi", 2, 45.00),
                ("Hotel - Midtown Suites (4 nights)", 4, 187.50),
                ("Meals & per diem allowance", 4, 38.00),
            ],
            "total": 1875.00,
        },
        {
            "filename": "receipt_software_licenses.pdf",
            "title": "SOFTWARE LICENSE RECEIPT",
            "vendor": "TechPro Solutions Inc.",
            "date": "2025-03-02",
            "category": "Software",
            "items": [
                ("Adobe Creative Cloud - Annual license", 1, 299.99),
                ("Slack Business subscription (5 seats)", 5, 12.50),
                ("GitHub Enterprise (monthly)", 1, 21.00),
                ("Zoom Pro annual plan", 1, 149.90),
                ("Figma Professional seat", 1, 90.00),
                ("1Password Teams (5 users)", 5, 3.99),
            ],
            "total": 625.00,
        },
        {
            "filename": "receipt_office_supplies.pdf",
            "title": "OFFICE SUPPLIES RECEIPT",
            "vendor": "Staples Business Solutions",
            "date": "2025-03-08",
            "category": "Office",
            "items": [
                ("Printer paper - A4 reams (5x)", 5, 8.99),
                ("Ballpoint pens - 12-pack", 3, 6.49),
                ("Sticky notes - assorted pack", 2, 4.99),
                ("File folders - 25-pack", 2, 9.49),
                ("Whiteboard markers set", 2, 7.49),
                ("Staples and binder clips", 1, 5.99),
                ("Notebooks - spiral bound (4x)", 4, 3.49),
                ("USB desk organizer", 1, 18.99),
            ],
            "total": 287.00,
        },
        {
            "filename": "receipt_food_catering.pdf",
            "title": "BUSINESS MEALS RECEIPT",
            "vendor": "Greenleaf Catering & Events",
            "date": "2025-03-12",
            "category": "Food",
            "items": [
                ("Quarterly team lunch (18 people)", 18, 18.50),
                ("Board meeting breakfast setup", 1, 85.00),
                ("Client dinner - The Steakhouse", 1, 156.00),
                ("Coffee & snacks weekly delivery", 4, 12.00),
            ],
            "total": 445.00,
        },
        {
            "filename": "receipt_marketing_campaign.pdf",
            "title": "MARKETING SERVICES RECEIPT",
            "vendor": "BrightSpark Digital Agency",
            "date": "2025-03-03",
            "category": "Marketing",
            "items": [
                ("Google Ads campaign management", 1, 450.00),
                ("LinkedIn sponsored posts (10 ads)", 10, 75.00),
                ("Email newsletter design & send", 2, 125.00),
                ("Social media graphics package", 1, 320.00),
                ("Trade show booth materials", 1, 180.00),
                ("Print brochures (500 units)", 500, 0.55),
                ("SEO optimization monthly fee", 1, 195.00),
            ],
            "total": 1820.00,
        },
        {
            "filename": "receipt_training_courses.pdf",
            "title": "EMPLOYEE TRAINING RECEIPT",
            "vendor": "ProfessionalEdge Training Center",
            "date": "2025-03-07",
            "category": "Training",
            "items": [
                ("Project Management Professional (PMP) course", 2, 149.00),
                ("Excel Advanced workshop (half-day)", 4, 35.00),
                ("Data Analysis with Python - online", 3, 49.00),
                ("Leadership & Communication seminar", 2, 89.00),
                ("First Aid & CPR certification", 2, 55.00),
            ],
            "total": 540.00,
        },
    ]

    for r in receipts:
        filepath = os.path.join(RECEIPTS_DIR, r["filename"])
        create_pdf_receipt(
            filepath=filepath,
            title=r["title"],
            vendor=r["vendor"],
            date=r["date"],
            items=r["items"],
            total=r["total"],
            category=r["category"],
        )

    print(f"All 6 PDF receipts created in {RECEIPTS_DIR}")


def create_budget_spreadsheet():
    """Create expense_budget.ods on Desktop with Budget column filled but Actual/Variance/Over_Budget empty."""
    try:
        from odf.opendocument import OpenDocumentSpreadsheet
        from odf.table import Table, TableRow, TableCell
        from odf.text import P
        from odf.style import Style, TableCellProperties, TextProperties, ParagraphProperties

        doc = OpenDocumentSpreadsheet()

        # Header style - bold
        header_style = Style(name="HeaderCell", family="table-cell")
        header_style.addElement(TextProperties(fontweight="bold"))
        doc.automaticstyles.addElement(header_style)

        # Number style for currency
        num_style = Style(name="NumericCell", family="table-cell")
        doc.automaticstyles.addElement(num_style)

        table = Table(name="Expenses")

        # Header row
        headers = ["Category", "Budget", "Actual", "Variance", "Over_Budget"]
        hrow = TableRow()
        for h in headers:
            cell = TableCell(stylename="HeaderCell", valuetype="string")
            cell.addElement(P(text=h))
            hrow.addElement(cell)
        table.addElement(hrow)

        # Data rows - Budget filled, Actual/Variance/Over_Budget EMPTY (task requires agent to fill)
        categories = [
            ("Travel",    2000),
            ("Software",   500),
            ("Office",     300),
            ("Food",       400),
            ("Marketing", 1500),
            ("Training",   600),
        ]

        for cat, budget in categories:
            row = TableRow()

            # Category (string)
            c_cat = TableCell(valuetype="string")
            c_cat.addElement(P(text=cat))
            row.addElement(c_cat)

            # Budget (numeric) - use 'value' keyword argument
            c_budget = TableCell(valuetype="float", value=str(budget))
            c_budget.addElement(P(text=str(budget)))
            row.addElement(c_budget)

            # Actual - EMPTY (agent must fill)
            row.addElement(TableCell())

            # Variance - EMPTY (agent must fill)
            row.addElement(TableCell())

            # Over_Budget - EMPTY (agent must fill)
            row.addElement(TableCell())

            table.addElement(row)

        doc.spreadsheet.addElement(table)

        # Ensure Desktop directory exists
        os.makedirs(f'{WORKDIR}/Desktop', exist_ok=True)
        doc.save(BUDGET_FILE)
        print(f"Budget spreadsheet created: {BUDGET_FILE}")

    except Exception as e:
        print(f"odfpy error: {e}. Falling back to openpyxl (.xlsx converted later).")
        # Fallback using openpyxl to create .xlsx then rename to .ods
        # LibreOffice will open .xlsx files; using .ods extension with xlsx content
        # is not ideal, but we'll try openpyxl for robustness
        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Expenses"

        headers = ["Category", "Budget", "Actual", "Variance", "Over_Budget"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)

        categories = [
            ("Travel",    2000),
            ("Software",   500),
            ("Office",     300),
            ("Food",       400),
            ("Marketing", 1500),
            ("Training",   600),
        ]

        for r, (cat, budget) in enumerate(categories, 2):
            ws.cell(row=r, column=1, value=cat)
            ws.cell(row=r, column=2, value=budget)
            # Columns 3 (Actual), 4 (Variance), 5 (Over_Budget) remain empty

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 14

        os.makedirs(f'{WORKDIR}/Desktop', exist_ok=True)
        wb.save(BUDGET_FILE)
        print(f"Budget spreadsheet created (xlsx as .ods fallback): {BUDGET_FILE}")


def create_initial():
    print("=== Creating initial state for osworld_multi_apps_doc_pdf_calc_008 ===")

    # 1. Create PDF receipts
    print("\n[1] Creating PDF receipts...")
    create_receipts()

    # 2. Create expense_budget.ods on Desktop
    print("\n[2] Creating expense_budget.ods on Desktop...")
    create_budget_spreadsheet()

    # 3. GUI-ready startup: open LibreOffice Calc with the budget spreadsheet
    #    and the file manager showing the receipts directory
    print("\n[3] Launching GUI apps...")
    launch_gui(f'libreoffice --calc "{BUDGET_FILE}"', delay_sec=2.5)
    launch_gui(f'nautilus "{RECEIPTS_DIR}"', delay_sec=1.5)

    print("\nGUI_READY: Launched LibreOffice Calc (budget) and Nautilus (receipts) with DISPLAY=:0")
    print(f"\nInitial files:")
    print(f"  Budget spreadsheet: {BUDGET_FILE}")
    print(f"  PDF receipts dir:   {RECEIPTS_DIR}")


create_initial()
