"""
Initial Setup: Create CRM deals spreadsheet with 200 rows of deal data
Task ID: calc_pivot_077
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_077'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'CRMDeals'

    # --- Headers ---
    headers = ['DealID', 'Company', 'SalesRep', 'Stage', 'DealSize', 'CloseDate']
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Data generation ---
    stages = ['Prospect', 'Qualified', 'Proposal', 'Negotiation', 'Closed Won', 'Closed Lost']

    companies = [
        'Apex Technologies', 'Bright Solutions', 'CloudNine Systems', 'DataStream Corp',
        'Echo Analytics', 'FusionWorks', 'GreenPath Labs', 'HorizonTech',
        'Innova Partners', 'JetForge Inc', 'Keystone Digital', 'Luminary Group',
        'MaverickAI', 'Nexus Dynamics', 'Optima Global', 'PrimeEdge',
        'Quantum Logic', 'RedShift Labs', 'Stellar Networks', 'TrueNorth Data',
        'Unified Systems', 'Vertex Solutions', 'WavePoint Inc', 'Xenon Enterprises',
        'Yield Analytics', 'Zenith Corp', 'Atlas Group', 'Beacon Digital',
        'Crestline Tech', 'DawnBreak Solutions', 'Elevate Inc', 'Frontier Labs',
        'GridPoint Systems', 'Harbor Analytics', 'IronClad Tech', 'Jupiter Data',
        'Kinetic Solutions', 'Lighthouse AI', 'Meridian Corp', 'NovaStrike'
    ]

    sales_reps = [
        'Sarah Chen', 'Marcus Johnson', 'Elena Rodriguez', 'David Kim',
        'Rachel Thompson', 'James Carter', 'Priya Patel', 'Michael O\'Brien',
        'Lisa Nakamura', 'Thomas Wright', 'Amanda Foster', 'Robert Singh'
    ]

    # Define the distribution to get exact ground truth values
    # Ground truth: Closed Won/Large=12, Prospect/Small=28, Grand total=200
    #
    # Size categories: Small (<5000), Medium (5000-20000), Large (>20000)
    # We'll define exact counts per (stage, size_category) cell
    #
    # Stage distribution (rows per stage):
    #   Prospect:     28 Small + 10 Medium + 5 Large  = 43
    #   Qualified:    8 Small  + 12 Medium + 6 Large  = 26
    #   Proposal:     6 Small  + 14 Medium + 8 Large  = 28
    #   Negotiation:  5 Small  + 10 Medium + 10 Large = 25
    #   Closed Won:   10 Small + 16 Medium + 12 Large = 38
    #   Closed Lost:  8 Small  + 18 Medium + 14 Large = 40
    #   Total:        65       + 80        + 55       = 200

    distribution = {
        'Prospect':    {'Small': 28, 'Medium': 10, 'Large': 5},
        'Qualified':   {'Small': 8,  'Medium': 12, 'Large': 6},
        'Proposal':    {'Small': 6,  'Medium': 14, 'Large': 8},
        'Negotiation': {'Small': 5,  'Medium': 10, 'Large': 10},
        'Closed Won':  {'Small': 10, 'Medium': 16, 'Large': 12},
        'Closed Lost': {'Small': 8,  'Medium': 18, 'Large': 14},
    }

    def random_deal_size(category):
        if category == 'Small':
            return random.randint(500, 4999)
        elif category == 'Medium':
            return random.randint(5000, 20000)
        else:  # Large
            return random.randint(20001, 50000)

    def random_date():
        year = random.choice([2024, 2025])
        month = random.randint(1, 12)
        day = random.randint(1, 28)
        return f'{year}-{month:02d}-{day:02d}'

    # Build all 200 rows
    rows = []
    deal_num = 1
    for stage in stages:
        for size_cat in ['Small', 'Medium', 'Large']:
            count = distribution[stage][size_cat]
            for _ in range(count):
                deal_id = f'D{deal_num:03d}'
                company = random.choice(companies)
                rep = random.choice(sales_reps)
                deal_size = random_deal_size(size_cat)
                close_date = random_date()
                rows.append([deal_id, company, rep, stage, deal_size, close_date])
                deal_num += 1

    # Shuffle rows so they aren't grouped by stage
    random.shuffle(rows)

    # Write data
    for r, row_data in enumerate(rows, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 5:  # DealSize column - number format
                cell.number_format = '#,##0'

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
