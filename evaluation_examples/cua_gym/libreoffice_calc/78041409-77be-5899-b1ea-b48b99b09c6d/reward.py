"""
Reward Script: Verify pivot table showing deal count by stage and deal size category
Task ID: calc_pivot_077
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20): PivotTable sheet exists
  Component 2 (0.25): Correct column headers (Small/Medium/Large + Grand Total)
  Component 3 (0.25): Correct row labels (all 6 stages + Grand Total)
  Component 4 (0.30): Correct data values (spot-check key cells and grand total)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_077'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: PivotTable sheet exists (0.20 points)
    # This is the primary task-introduced change: a new sheet must be created.
    # Initial env has only 'CRMDeals'; golden env has 'CRMDeals' + 'PivotTable'.
    try:
        pivot_ws = None
        for sn in wb.sheetnames:
            if 'pivot' in sn.lower():
                pivot_ws = wb[sn]
                break
        if pivot_ws is not None:
            print(f"PASS: Component 1 -- PivotTable sheet found: '{pivot_ws.title}' (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 -- No sheet containing 'pivot' in name. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0  # No pivot sheet means nothing else to check
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 2: Correct column headers - deal size categories (0.25 points)
    # The pivot should have columns: Small (<5000), Medium (5000-20000), Large (>20000), Grand Total
    try:
        # Find the header row - look for a row containing 'Small' or 'Medium' or 'Large'
        header_row = None
        for row_idx in range(1, min(pivot_ws.max_row + 1, 10)):
            row_vals = []
            for col_idx in range(1, pivot_ws.max_column + 1):
                v = pivot_ws.cell(row=row_idx, column=col_idx).value
                if v is not None:
                    row_vals.append(str(v).lower())
            row_text = ' '.join(row_vals)
            if 'small' in row_text and ('medium' in row_text or 'large' in row_text):
                header_row = row_idx
                break

        if header_row is not None:
            headers = []
            for col_idx in range(1, pivot_ws.max_column + 1):
                v = pivot_ws.cell(row=header_row, column=col_idx).value
                if v is not None:
                    headers.append(str(v))

            has_small = any('small' in h.lower() for h in headers)
            has_medium = any('medium' in h.lower() for h in headers)
            has_large = any('large' in h.lower() for h in headers)

            matched = sum([has_small, has_medium, has_large])
            if matched == 3:
                print(f"PASS: Component 2 -- All 3 size categories found in headers: {headers} (0.25 pts)")
                total_score += 0.25
            elif matched >= 2:
                partial = round(0.25 * matched / 3, 2)
                print(f"PARTIAL: Component 2 -- {matched}/3 categories found: {headers} ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 2 -- Only {matched}/3 categories. Headers: {headers}")
        else:
            print(f"FAIL: Component 2 -- Could not find header row with size categories")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Correct row labels - all 6 stages present (0.25 points)
    # Expected stages: Prospect, Qualified, Proposal, Negotiation, Closed Won, Closed Lost
    try:
        expected_stages = {'prospect', 'qualified', 'proposal', 'negotiation', 'closed won', 'closed lost'}
        found_stages = set()

        # Scan column A for stage names (below the header row)
        data_start = (header_row + 1) if header_row else 2
        for row_idx in range(data_start, pivot_ws.max_row + 1):
            v = pivot_ws.cell(row=row_idx, column=1).value
            if v is not None:
                v_lower = str(v).strip().lower()
                if v_lower in expected_stages:
                    found_stages.add(v_lower)

        matched_stages = len(found_stages)
        if matched_stages == 6:
            print(f"PASS: Component 3 -- All 6 stages found: {found_stages} (0.25 pts)")
            total_score += 0.25
        elif matched_stages >= 4:
            partial = round(0.25 * matched_stages / 6, 2)
            print(f"PARTIAL: Component 3 -- {matched_stages}/6 stages found: {found_stages} ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 -- Only {matched_stages}/6 stages found: {found_stages}")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Correct data values (0.30 points)
    # Verify key ground-truth values from context:
    #   - Prospect/Small = 28
    #   - Closed Won/Large = 12
    #   - Grand Total = 200
    # We need to locate cells by matching row/column labels dynamically.
    try:
        checks_passed = 0
        total_checks = 3

        if header_row is None:
            print(f"FAIL: Component 4 -- Cannot verify values without header row")
        else:
            # Build column mapping: category name -> column index
            col_map = {}
            for col_idx in range(1, pivot_ws.max_column + 1):
                v = pivot_ws.cell(row=header_row, column=col_idx).value
                if v is not None:
                    col_map[str(v).strip().lower()] = col_idx

            # Build row mapping: stage name -> row index
            row_map = {}
            for row_idx in range(data_start, pivot_ws.max_row + 1):
                v = pivot_ws.cell(row=row_idx, column=1).value
                if v is not None:
                    row_map[str(v).strip().lower()] = row_idx

            # Check 1: Prospect/Small = 28
            prospect_row = row_map.get('prospect')
            small_col = None
            for k, v in col_map.items():
                if 'small' in k:
                    small_col = v
                    break
            if prospect_row and small_col:
                val = pivot_ws.cell(row=prospect_row, column=small_col).value
                if val is not None and int(val) == 28:
                    print(f"PASS: Check 4a -- Prospect/Small = {val} (expected 28)")
                    checks_passed += 1
                else:
                    print(f"FAIL: Check 4a -- Prospect/Small = {val} (expected 28)")
            else:
                print(f"FAIL: Check 4a -- Could not locate Prospect row or Small column")

            # Check 2: Closed Won/Large = 12
            closed_won_row = row_map.get('closed won')
            large_col = None
            for k, v in col_map.items():
                if 'large' in k:
                    large_col = v
                    break
            if closed_won_row and large_col:
                val = pivot_ws.cell(row=closed_won_row, column=large_col).value
                if val is not None and int(val) == 12:
                    print(f"PASS: Check 4b -- Closed Won/Large = {val} (expected 12)")
                    checks_passed += 1
                else:
                    print(f"FAIL: Check 4b -- Closed Won/Large = {val} (expected 12)")
            else:
                print(f"FAIL: Check 4b -- Could not locate Closed Won row or Large column")

            # Check 3: Grand Total = 200
            grand_total_row = row_map.get('grand total')
            grand_total_col = None
            for k, v in col_map.items():
                if 'grand' in k and 'total' in k:
                    grand_total_col = v
                    break
            # Also check the last populated column in grand total row
            if grand_total_row:
                if grand_total_col:
                    val = pivot_ws.cell(row=grand_total_row, column=grand_total_col).value
                else:
                    # Try last column
                    val = pivot_ws.cell(row=grand_total_row, column=pivot_ws.max_column).value
                if val is not None and int(val) == 200:
                    print(f"PASS: Check 4c -- Grand Total = {val} (expected 200)")
                    checks_passed += 1
                else:
                    print(f"FAIL: Check 4c -- Grand Total = {val} (expected 200)")
            else:
                # Try to find grand total in any row
                for row_idx in range(data_start, pivot_ws.max_row + 1):
                    v = pivot_ws.cell(row=row_idx, column=1).value
                    if v and 'total' in str(v).lower():
                        last_col_val = pivot_ws.cell(row=row_idx, column=pivot_ws.max_column).value
                        if last_col_val is not None and int(last_col_val) == 200:
                            print(f"PASS: Check 4c -- Grand Total = {last_col_val} (expected 200)")
                            checks_passed += 1
                            break
                else:
                    print(f"FAIL: Check 4c -- Could not locate Grand Total row")

            score_4 = round(0.30 * checks_passed / total_checks, 2)
            if checks_passed == total_checks:
                print(f"PASS: Component 4 -- All {total_checks} value checks passed ({score_4} pts)")
                total_score += score_4
            elif checks_passed > 0:
                print(f"PARTIAL: Component 4 -- {checks_passed}/{total_checks} value checks passed ({score_4} pts)")
                total_score += score_4
            else:
                print(f"FAIL: Component 4 -- 0/{total_checks} value checks passed")

    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
