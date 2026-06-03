"""
Reward Script: Delete 'Old Version' sheet and verify no broken cross-sheet formulas
Task ID: calc_gsi_065
Domain: libreoffice_calc
Scoring:
  Component 1: 'Old Version' sheet is absent (0.5 points)
  Component 2: No #REF! errors in any remaining sheet (0.3 points)
  Component 3: Cross-sheet formulas still reference valid sheets (0.2 points)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_065'


def persist_app_state(domain: str):
    """Try to save any open LibreOffice document before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    sheet_names = wb.sheetnames
    print(f"INFO: Found sheets: {sheet_names}")

    # Component 1: 'Old Version' sheet is absent (0.5 points)
    # This is the primary task action — the sheet must be deleted.
    # In initial_env, 'Old Version' exists, so this check FAILS -> 0 points.
    # In golden_env, 'Old Version' is deleted, so this check PASSES -> 0.5 points.
    try:
        if 'Old Version' not in sheet_names:
            print(f"PASS: Component 1 — 'Old Version' sheet is absent (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — 'Old Version' sheet still exists")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: No #REF! errors in any remaining sheet (0.3 points)
    # After deleting 'Old Version', formulas referencing other sheets must not be broken.
    # In initial_env, 'Old Version' still exists so this component is gated by Component 1.
    # We only award points if 'Old Version' is absent AND no #REF! errors exist.
    try:
        if 'Old Version' in sheet_names:
            # Gate: sheet not deleted yet, so this component does not apply
            print(f"FAIL: Component 2 — Gated: 'Old Version' still present, cannot verify post-deletion integrity")
        else:
            ref_errors = []
            for name in sheet_names:
                ws = wb[name]
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                    for cell in row:
                        val = cell.value
                        if val is not None and isinstance(val, str) and '#REF' in val.upper():
                            ref_errors.append(f"{name}!{cell.coordinate}: {val}")

            if len(ref_errors) == 0:
                print(f"PASS: Component 2 — No #REF! errors in any remaining sheet (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — Found {len(ref_errors)} #REF! error(s):")
                for err in ref_errors:
                    print(f"  {err}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Cross-sheet formulas reference valid sheets only (0.2 points)
    # After deletion, formulas that reference other sheets should still point to existing sheets.
    # Gated on 'Old Version' being absent (same logic as Component 2).
    try:
        if 'Old Version' in sheet_names:
            print(f"FAIL: Component 3 — Gated: 'Old Version' still present")
        else:
            # Collect all formulas and check that any sheet references point to existing sheets
            broken_refs = []
            valid_sheet_names_lower = {s.lower() for s in sheet_names}

            for name in sheet_names:
                ws = wb[name]
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                    for cell in row:
                        val = cell.value
                        if val is not None and isinstance(val, str) and val.startswith('='):
                            # Extract sheet references from formula
                            # Patterns: 'Sheet Name'!A1 or SheetName!A1
                            refs = re.findall(r"'([^']+)'!", val)
                            refs += re.findall(r"(?<!=)(?<!['\w])([A-Za-z_]\w+)!", val)
                            for ref_sheet in refs:
                                if ref_sheet.lower() not in valid_sheet_names_lower:
                                    broken_refs.append(f"{name}!{cell.coordinate}: references non-existent sheet '{ref_sheet}' in formula '{val}'")

            if len(broken_refs) == 0:
                print(f"PASS: Component 3 — All cross-sheet formula references are valid (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 3 — Found {len(broken_refs)} broken sheet reference(s):")
                for br in broken_refs:
                    print(f"  {br}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
