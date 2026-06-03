"""
Initial Setup: Delete 'Old Version' sheet and verify no #REF errors
Task ID: calc_gsi_065
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_065'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Styles ---
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    currency_fmt = '$#,##0.00'
    date_fmt = 'yyyy-mm-dd'

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def style_data_row(ws, row, cols):
        for c in range(1, cols + 1):
            ws.cell(row=row, column=c).border = thin_border

    # ========== Sheet 1: Employee Data ==========
    ws1 = wb.active
    ws1.title = "Employee Data"
    headers1 = ["Employee ID", "Name", "Department", "Position", "Salary", "Start Date", "Status"]
    for c, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=c, value=h)
    style_header(ws1, 1, len(headers1))

    employees = [
        ["EMP-001", "Sarah Chen", "Engineering", "Senior Developer", 95000, "2021-03-15", "Active"],
        ["EMP-002", "Marcus Johnson", "Marketing", "Campaign Manager", 72000, "2022-06-01", "Active"],
        ["EMP-003", "Priya Patel", "Engineering", "Tech Lead", 115000, "2020-01-10", "Active"],
        ["EMP-004", "James Wilson", "Finance", "Financial Analyst", 68000, "2023-02-20", "Active"],
        ["EMP-005", "Elena Rodriguez", "HR", "Recruiter", 62000, "2022-11-05", "Active"],
        ["EMP-006", "David Kim", "Engineering", "Junior Developer", 65000, "2024-01-08", "Active"],
        ["EMP-007", "Rachel Thompson", "Marketing", "Content Strategist", 70000, "2021-09-12", "Active"],
        ["EMP-008", "Omar Hassan", "Finance", "Controller", 98000, "2019-07-22", "Active"],
        ["EMP-009", "Lisa Wang", "Engineering", "QA Engineer", 78000, "2022-04-18", "Active"],
        ["EMP-010", "Michael Brown", "HR", "HR Director", 105000, "2018-05-30", "Active"],
        ["EMP-011", "Anna Kowalski", "Marketing", "Digital Analyst", 67000, "2023-08-14", "Active"],
        ["EMP-012", "Carlos Mendez", "Finance", "Payroll Specialist", 58000, "2024-03-01", "Active"],
    ]
    for r, row_data in enumerate(employees, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            if c == 5:
                cell.number_format = currency_fmt
            elif c == 6:
                cell.number_format = date_fmt
        style_data_row(ws1, r, len(headers1))

    # Summary row
    ws1.cell(row=14, column=4, value="Total Salaries:")
    ws1.cell(row=14, column=4).font = Font(bold=True)
    ws1.cell(row=14, column=5, value="=SUM(E2:E13)")
    ws1.cell(row=14, column=5).number_format = currency_fmt
    ws1.cell(row=14, column=5).font = Font(bold=True)

    ws1.cell(row=15, column=4, value="Average Salary:")
    ws1.cell(row=15, column=4).font = Font(bold=True)
    ws1.cell(row=15, column=5, value="=AVERAGE(E2:E13)")
    ws1.cell(row=15, column=5).number_format = currency_fmt
    ws1.cell(row=15, column=5).font = Font(bold=True)

    ws1.cell(row=16, column=4, value="Employee Count:")
    ws1.cell(row=16, column=4).font = Font(bold=True)
    ws1.cell(row=16, column=5, value="=COUNTA(A2:A13)")
    ws1.cell(row=16, column=5).font = Font(bold=True)

    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 14
    ws1.column_dimensions["D"].width = 20
    ws1.column_dimensions["E"].width = 14
    ws1.column_dimensions["F"].width = 14
    ws1.column_dimensions["G"].width = 10

    # ========== Sheet 2: Department Budget ==========
    ws2 = wb.create_sheet("Department Budget")
    headers2 = ["Department", "Q1 Budget", "Q2 Budget", "Q3 Budget", "Q4 Budget", "Annual Total", "Headcount"]
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=c, value=h)
    style_header(ws2, 1, len(headers2))

    budgets = [
        ["Engineering", 450000, 475000, 500000, 520000],
        ["Marketing", 180000, 195000, 210000, 200000],
        ["Finance", 120000, 125000, 130000, 135000],
        ["HR", 95000, 100000, 105000, 110000],
    ]
    for r, row_data in enumerate(budgets, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if c >= 2:
                cell.number_format = currency_fmt
        # Annual Total formula
        ws2.cell(row=r, column=6, value=f"=SUM(B{r}:E{r})")
        ws2.cell(row=r, column=6).number_format = currency_fmt
        # Headcount formula referencing Employee Data (cross-sheet)
        dept_name = row_data[0]
        ws2.cell(row=r, column=7, value=f'=COUNTIF(\'Employee Data\'!C:C,"{dept_name}")')
        style_data_row(ws2, r, len(headers2))

    # Totals row
    ws2.cell(row=6, column=1, value="Total").font = Font(bold=True)
    for c in range(2, 7):
        col_letter = openpyxl.utils.get_column_letter(c)
        ws2.cell(row=6, column=c, value=f"=SUM({col_letter}2:{col_letter}5)")
        ws2.cell(row=6, column=c).number_format = currency_fmt
        ws2.cell(row=6, column=c).font = Font(bold=True)
    ws2.cell(row=6, column=7, value="=SUM(G2:G5)")
    ws2.cell(row=6, column=7).font = Font(bold=True)
    style_data_row(ws2, 6, len(headers2))

    ws2.column_dimensions["A"].width = 14
    for col in ["B", "C", "D", "E", "F"]:
        ws2.column_dimensions[col].width = 14
    ws2.column_dimensions["G"].width = 12

    # ========== Sheet 3: Old Version (to be deleted by agent) ==========
    ws3 = wb.create_sheet("Old Version")
    headers3 = ["Employee ID", "Name", "Department", "Salary", "Hire Date"]
    for c, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=c, value=h)
    style_header(ws3, 1, len(headers3))

    old_data = [
        ["EMP-001", "Sarah Chen", "Engineering", 88000, "2021-03-15"],
        ["EMP-002", "Marcus Johnson", "Marketing", 68000, "2022-06-01"],
        ["EMP-003", "Priya Patel", "Engineering", 108000, "2020-01-10"],
        ["EMP-004", "James Wilson", "Finance", 65000, "2023-02-20"],
        ["EMP-005", "Elena Rodriguez", "HR", 59000, "2022-11-05"],
        ["EMP-006", "David Kim", "Engineering", 60000, "2024-01-08"],
        ["EMP-007", "Rachel Thompson", "Marketing", 66000, "2021-09-12"],
        ["EMP-008", "Omar Hassan", "Finance", 92000, "2019-07-22"],
    ]
    for r, row_data in enumerate(old_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            if c == 4:
                cell.number_format = currency_fmt
            elif c == 5:
                cell.number_format = date_fmt
        style_data_row(ws3, r, len(headers3))

    # Mark it as outdated
    ws3.cell(row=10, column=1, value="Note: This data is from the previous fiscal year and is no longer current.")
    ws3.cell(row=10, column=1).font = Font(italic=True, color="999999")
    ws3.merge_cells("A10:E10")

    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 14
    ws3.column_dimensions["D"].width = 14
    ws3.column_dimensions["E"].width = 14

    # ========== Sheet 4: Summary ==========
    ws4 = wb.create_sheet("Summary")
    ws4.cell(row=1, column=1, value="Company Workforce Summary")
    ws4.cell(row=1, column=1).font = Font(name="Arial", size=14, bold=True)
    ws4.merge_cells("A1:D1")

    summary_headers = ["Metric", "Value"]
    for c, h in enumerate(summary_headers, 1):
        ws4.cell(row=3, column=c, value=h)
    style_header(ws4, 3, len(summary_headers))

    # Formulas referencing Employee Data and Department Budget (NOT Old Version)
    metrics = [
        ["Total Employees", "=COUNTA('Employee Data'!A2:A13)"],
        ["Total Salary Expense", "=SUM('Employee Data'!E2:E13)"],
        ["Average Salary", "=AVERAGE('Employee Data'!E2:E13)"],
        ["Highest Salary", "=MAX('Employee Data'!E2:E13)"],
        ["Lowest Salary", "=MIN('Employee Data'!E2:E13)"],
        ["Total Annual Budget", "='Department Budget'!F6"],
        ["Number of Departments", "=COUNTA('Department Budget'!A2:A5)"],
    ]
    for r, (label, formula) in enumerate(metrics, 4):
        ws4.cell(row=r, column=1, value=label)
        ws4.cell(row=r, column=1).font = Font(bold=True)
        ws4.cell(row=r, column=2, value=formula)
        if "Salary" in label or "Budget" in label:
            ws4.cell(row=r, column=2).number_format = currency_fmt
        style_data_row(ws4, r, 2)

    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
