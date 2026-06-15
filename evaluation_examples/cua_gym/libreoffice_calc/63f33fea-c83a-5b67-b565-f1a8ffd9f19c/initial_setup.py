"""
Initial Setup: Build a project RACI matrix structure (empty, no assignments)
Task ID: calc_gpm_072
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_072'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

TEAL_ARGB = 'FF006666'       # dark teal RGB(0,102,102)
WHITE_ARGB = 'FFFFFFFF'

def launch_gui(command: str, delay_sec: float = 1.0):
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'RACI'

    thin = Side(style='thin', color='000000')
    all_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Title row: merge A1:H1 ---
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = 'Project RACI Matrix - CRM Implementation'
    title_cell.font = Font(size=14, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color=TEAL_ARGB, end_color=TEAL_ARGB, fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Row 3: Headers ---
    headers = ['Deliverable', 'Project Mgr', 'Dev Lead', 'QA Lead',
               'Business Analyst', 'Stakeholder', 'DBA', 'UX Designer']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color=TEAL_ARGB, end_color=TEAL_ARGB, fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = all_border

    # --- Rows 4-15: Deliverable names only (NO RACI values) ---
    deliverables = [
        'Requirements Document',
        'System Architecture',
        'Database Design',
        'UI Wireframes',
        'API Development',
        'Unit Tests',
        'Integration Tests',
        'Data Migration',
        'User Training',
        'Documentation',
        'UAT',
        'Go-Live Approval',
    ]

    for r_idx, name in enumerate(deliverables, 4):
        cell = ws.cell(row=r_idx, column=1, value=name)
        cell.border = all_border
        # Add borders to empty RACI cells too
        for c in range(2, 9):
            ws.cell(row=r_idx, column=c).border = all_border

    # --- Column widths ---
    ws.column_dimensions['A'].width = 22
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col_letter].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
