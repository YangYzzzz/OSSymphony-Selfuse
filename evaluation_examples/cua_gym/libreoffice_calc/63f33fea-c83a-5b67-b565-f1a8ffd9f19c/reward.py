"""
Reward Script: Build a project RACI matrix with role assignments and validation-controlled entries.
Task ID: calc_gpm_072
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): RACI values populated in B4:H15 with valid assignment rules
  Component 2 (0.20): Data validation (dropdown list) applied to B4:H15
  Component 3 (0.20): Conditional formatting rules for R/A/C/I colors on B4:H15
  Component 4 (0.20): COUNTIF formulas in rows 18-21 with Role Summary label
  Component 5 (0.15): Legend at A23:B26 with matching fill colors
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_072'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: RACI sheet must exist
    if 'RACI' not in wb.sheetnames:
        print("FAIL: 'RACI' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['RACI']

    # ========================================================================
    # Component 1: RACI values populated in B4:H15 with valid assignment rules
    # (0.25 points)
    # This FAILS on initial (no values) and PASSES on golden (values present).
    # ========================================================================
    try:
        valid_vals = {'R', 'A', 'C', 'I', '', None}
        all_valid = True
        rows_with_one_a = 0
        rows_with_at_least_one_r = 0
        total_raci_cells_filled = 0

        for row_idx in range(4, 16):  # rows 4-15
            row_vals = []
            for col_idx in range(2, 9):  # cols B-H
                val = ws.cell(row=row_idx, column=col_idx).value
                row_vals.append(val)
                if val is not None and str(val).strip() != '':
                    total_raci_cells_filled += 1
                if val is not None and str(val).strip() not in {'R', 'A', 'C', 'I', ''}:
                    all_valid = False

            # Check exactly one A per row
            a_count = sum(1 for v in row_vals if v is not None and str(v).strip() == 'A')
            r_count = sum(1 for v in row_vals if v is not None and str(v).strip() == 'R')
            if a_count == 1:
                rows_with_one_a += 1
            if r_count >= 1:
                rows_with_at_least_one_r += 1

        # Need at least some RACI cells filled (initial has 0)
        if total_raci_cells_filled >= 60 and all_valid and rows_with_one_a >= 10 and rows_with_at_least_one_r >= 10:
            print(f"PASS: Component 1 - RACI values populated ({total_raci_cells_filled} cells), "
                  f"{rows_with_one_a}/12 rows with exactly 1 A, "
                  f"{rows_with_at_least_one_r}/12 rows with at least 1 R (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 - RACI values: {total_raci_cells_filled} filled cells, "
                  f"all_valid={all_valid}, rows_1A={rows_with_one_a}, rows_R={rows_with_at_least_one_r}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # ========================================================================
    # Component 2: Data validation (dropdown list) applied to B4:H15
    # (0.20 points)
    # Initial has NO data validations; golden has list validation on B4:H15.
    # ========================================================================
    try:
        has_list_validation = False
        for dv in ws.data_validations.dataValidation:
            if dv.type == 'list':
                # Check that the validation covers the RACI range B4:H15
                dv_range_str = str(dv.sqref)
                # Check formula contains R,A,C,I
                formula_str = str(dv.formula1) if dv.formula1 else ''
                has_raci = all(letter in formula_str for letter in ['R', 'A', 'C', 'I'])
                if has_raci:
                    has_list_validation = True
                    break

        if has_list_validation:
            print(f"PASS: Component 2 - Data validation with R/A/C/I list found (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 - No list data validation with R/A/C/I found")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # ========================================================================
    # Component 3: Conditional formatting rules for R/A/C/I on B4:H15
    # (0.20 points)
    # Initial has NO conditional formatting; golden has 4 rules + data bar.
    # ========================================================================
    try:
        cf_rules_found = set()
        has_data_bar = False

        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            for rule in cf.rules:
                if rule.type == 'cellIs' and hasattr(rule, 'formula') and rule.formula:
                    formula_val = rule.formula[0].strip().strip('"').strip("'")
                    if formula_val in ('R', 'A', 'C', 'I'):
                        cf_rules_found.add(formula_val)
                elif rule.type == 'dataBar':
                    has_data_bar = True

        # Need at least 3 of the 4 RACI conditional formatting rules
        cf_score = 0.0
        if len(cf_rules_found) >= 4:
            cf_score = 0.15
            print(f"PASS: Component 3a - All 4 RACI conditional formatting rules found (0.15 pts)")
        elif len(cf_rules_found) >= 2:
            cf_score = 0.08
            print(f"PARTIAL: Component 3a - {len(cf_rules_found)}/4 RACI CF rules found (0.08 pts)")
        else:
            print(f"FAIL: Component 3a - Only {len(cf_rules_found)}/4 RACI CF rules found")

        if has_data_bar:
            cf_score += 0.05
            print(f"PASS: Component 3b - Data bar found on R-count row (0.05 pts)")
        else:
            print(f"FAIL: Component 3b - No data bar found")

        total_score += cf_score

    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # ========================================================================
    # Component 4: COUNTIF formulas in rows 18-21 + Role Summary label
    # (0.20 points)
    # Initial has NONE of these; golden has "Role Summary" at A17 and COUNTIF
    # formulas across B18:H21.
    # ========================================================================
    try:
        comp4_score = 0.0

        # Check "Role Summary" label at A17
        a17_val = ws['A17'].value
        if a17_val and 'role summary' in str(a17_val).lower():
            comp4_score += 0.04
            print(f"PASS: Component 4a - 'Role Summary' label found at A17 (0.04 pts)")
        else:
            print(f"FAIL: Component 4a - A17 value: {repr(a17_val)}")

        # Check COUNTIF formulas in rows 18-21
        countif_found = 0
        expected_letters = {'R': 18, 'A': 19, 'C': 20, 'I': 21}

        for letter, row_idx in expected_letters.items():
            for col_idx in range(2, 9):  # B through H
                val = ws.cell(row=row_idx, column=col_idx).value
                if val and isinstance(val, str) and 'COUNTIF' in val.upper() and f'"{letter}"' in val:
                    countif_found += 1

        # Expect 7 columns x 4 letters = 28 COUNTIF formulas
        if countif_found >= 24:
            comp4_score += 0.16
            print(f"PASS: Component 4b - {countif_found}/28 COUNTIF formulas found (0.16 pts)")
        elif countif_found >= 14:
            comp4_score += 0.08
            print(f"PARTIAL: Component 4b - {countif_found}/28 COUNTIF formulas found (0.08 pts)")
        else:
            print(f"FAIL: Component 4b - Only {countif_found}/28 COUNTIF formulas found")

        total_score += comp4_score

    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # ========================================================================
    # Component 5: Legend at A23:B26 with matching fill colors
    # (0.15 points)
    # Initial has NO legend; golden has R/A/C/I legend with colored fills.
    # ========================================================================
    try:
        legend_entries = {
            23: ('R', 'Responsible'),
            24: ('A', 'Accountable'),
            25: ('C', 'Consulted'),
            26: ('I', 'Informed'),
        }

        legend_found = 0
        legend_with_color = 0

        for row_idx, (expected_key, expected_desc) in legend_entries.items():
            a_val = ws.cell(row=row_idx, column=1).value
            b_val = ws.cell(row=row_idx, column=2).value

            if a_val and str(a_val).strip() == expected_key and b_val and expected_desc.lower() in str(b_val).lower():
                legend_found += 1

                # Check if A-column cell has a non-default fill color
                try:
                    fill_rgb = ws.cell(row=row_idx, column=1).fill.fgColor.rgb
                    if fill_rgb and fill_rgb != '00000000' and fill_rgb != '0':
                        legend_with_color += 1
                except:
                    pass

        comp5_score = 0.0
        if legend_found >= 4 and legend_with_color >= 3:
            comp5_score = 0.15
            print(f"PASS: Component 5 - Legend complete: {legend_found}/4 entries, {legend_with_color}/4 colored (0.15 pts)")
        elif legend_found >= 2:
            comp5_score = 0.07
            print(f"PARTIAL: Component 5 - Legend partial: {legend_found}/4 entries, {legend_with_color}/4 colored (0.07 pts)")
        else:
            print(f"FAIL: Component 5 - Legend: {legend_found}/4 entries found")

        total_score += comp5_score

    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
