"""
Initial Setup: HR export with hierarchical blank fill-down pattern
Task ID: osworld_calc_fill_blanks_above_007
Domain: libreoffice_calc

Creates a spreadsheet with Division (A) and Department (B) columns
that use the fill-down convention (blanks where the same group continues).
The agent must fill those blanks and produce a summary count table.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_fill_blanks_above_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: HR Data ---
    ws = wb.active
    ws.title = "HR Data"

    # Headers
    headers = ["Division", "Department", "Employee Name", "Salary"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # HR data with hierarchical blank fill-down pattern.
    # Division column: filled only on the first row of each Division block.
    # Department column: filled only on the first row of each Dept block.
    # fmt: off
    data = [
        # Division,       Department,   Employee Name,          Salary
        ("Technology",    "Engineering", "Priya Sharma",         95000),
        (None,            None,          "Liam O'Brien",         87500),
        (None,            None,          "Mei-Ling Zhou",        91200),
        (None,            "QA",          "Carlos Reyes",         74000),
        (None,            None,          "Fatima Al-Hassan",     71500),
        (None,            "IT Support",  "Derek Novak",          68000),
        (None,            None,          "Ananya Patel",         66800),
        ("Operations",    "Logistics",   "Samuel Abebe",         72000),
        (None,            None,          "Hannah Bergstrom",     69500),
        (None,            None,          "Rico Delgado",         70200),
        (None,            "Facilities",  "Yuki Tanaka",          63000),
        (None,            None,          "Oluwaseun Adeyemi",    61500),
        ("Finance",       "Accounting",  "Eleanor Whitfield",    84000),
        (None,            None,          "James Okonkwo",        81000),
        (None,            "Audit",       "Sophia Lindqvist",     88500),
        (None,            None,          "Tariq Mansoor",        86000),
        (None,            None,          "Grace Nwosu",          83500),
        ("HR",            "Recruitment", "Nathan Kowalski",      76000),
        (None,            None,          "Isabelle Moreau",      74500),
        (None,            "Training",    "Ahmed El-Sayed",       72000),
    ]
    # fmt: on

    for r, (division, department, name, salary) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=division)
        ws.cell(row=r, column=2, value=department)
        ws.cell(row=r, column=3, value=name)
        ws.cell(row=r, column=4, value=salary)

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 14

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
