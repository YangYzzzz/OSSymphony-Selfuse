"""
Reward Script: Fill blank cells in Division and Department columns (fill-down pattern),
               then count employees per Division-Department combination.
Task ID: osworld_calc_fill_blanks_above_007
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): Division column (A) has NO blank (None) cells — all rows filled
  Component 2 (0.35): Department column (B) has NO blank (None) cells — all rows filled
                       AND all values match expected fill-down pattern
  Component 3 (0.30): Summary table exists with correct Division-Department-EmployeeCount
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_fill_blanks_above_007'

# Ground truth: expected fill-down values for Division (col A) and Department (col B)
# Row 1 = header. Data rows are rows 2-21 (20 data rows).
EXPECTED_DIVISION = [
    'Technology', 'Technology', 'Technology',   # rows 2-4 (Engineering)
    'Technology', 'Technology',                  # rows 5-6 (QA)
    'Technology', 'Technology',                  # rows 7-8 (IT Support)
    'Operations', 'Operations', 'Operations',    # rows 9-11 (Logistics)
    'Operations', 'Operations',                  # rows 12-13 (Facilities)
    'Finance', 'Finance',                        # rows 14-15 (Accounting)
    'Finance', 'Finance', 'Finance',             # rows 16-18 (Audit)
    'HR', 'HR',                                  # rows 19-20 (Recruitment)
    'HR',                                        # row 21 (Training)
]

EXPECTED_DEPARTMENT = [
    'Engineering', 'Engineering', 'Engineering', # rows 2-4
    'QA', 'QA',                                  # rows 5-6
    'IT Support', 'IT Support',                  # rows 7-8
    'Logistics', 'Logistics', 'Logistics',       # rows 9-11
    'Facilities', 'Facilities',                  # rows 12-13
    'Accounting', 'Accounting',                  # rows 14-15
    'Audit', 'Audit', 'Audit',                   # rows 16-18
    'Recruitment', 'Recruitment',                # rows 19-20
    'Training',                                  # row 21
]

# Ground truth for summary table: Division-Department -> Employee Count
EXPECTED_SUMMARY = {
    ('Finance', 'Accounting'): 2,
    ('Finance', 'Audit'): 3,
    ('HR', 'Recruitment'): 2,
    ('HR', 'Training'): 1,
    ('Operations', 'Facilities'): 2,
    ('Operations', 'Logistics'): 3,
    ('Technology', 'Engineering'): 3,
    ('Technology', 'IT Support'): 2,
    ('Technology', 'QA'): 2,
}

# Number of data rows in initial_env that were originally blank in each column
# Division (col A): 16 blanks originally; Department (col B): 11 blanks originally
# The task is complete ONLY when all blanks are filled AND values are correct.
INITIAL_BLANK_ROWS_DIV = {3, 4, 5, 6, 7, 8, 10, 11, 12, 13, 15, 16, 17, 18, 20, 21}
INITIAL_BLANK_ROWS_DEPT = {3, 4, 6, 8, 10, 11, 13, 15, 17, 18, 20}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws_hr = None
    if 'HR Data' in wb.sheetnames:
        ws_hr = wb['HR Data']

    # ----------------------------------------------------------------
    # Component 1: Division column (A) blanks filled correctly (0.35 pts)
    # Checks ONLY the rows that were originally blank in the initial_env.
    # These rows must now be filled with the correct value from above.
    # This component FAILS on initial_env (those cells are None there)
    # and PASSES on golden_env (those cells are filled).
    # ----------------------------------------------------------------
    try:
        if ws_hr is None:
            print("FAIL: Component 1 — 'HR Data' sheet not found")
        else:
            # Only check the rows that were originally blank in Division col
            originally_blank_correct = 0
            total_originally_blank = len(INITIAL_BLANK_ROWS_DIV)
            mismatches = []

            for i, expected_div in enumerate(EXPECTED_DIVISION):
                row_idx = i + 2  # rows 2-21
                if row_idx not in INITIAL_BLANK_ROWS_DIV:
                    continue  # Skip rows that were already filled (precondition)
                actual = ws_hr.cell(row=row_idx, column=1).value
                if actual is not None and str(actual).strip() == expected_div:
                    originally_blank_correct += 1
                else:
                    mismatches.append(f"A{row_idx}: expected '{expected_div}', got '{actual}'")

            if mismatches:
                for m in mismatches[:5]:  # Show first 5 mismatches
                    print(f"  FAIL: Division {m}")
                if len(mismatches) > 5:
                    print(f"  ... and {len(mismatches) - 5} more failures")

            if originally_blank_correct == total_originally_blank:
                print(f"PASS: Component 1 — All {total_originally_blank} originally-blank Division cells filled correctly (0.35 pts)")
                total_score += 0.35
            elif originally_blank_correct > 0:
                partial = round(0.35 * (originally_blank_correct / total_originally_blank), 4)
                print(f"PARTIAL: Component 1 — {originally_blank_correct}/{total_originally_blank} originally-blank Division cells filled correctly ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 1 — No originally-blank Division cells were filled correctly (0/{total_originally_blank})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ----------------------------------------------------------------
    # Component 2: Department column (B) blanks filled correctly (0.35 pts)
    # Checks ONLY the rows that were originally blank in the initial_env.
    # These rows must now be filled with the correct value from above.
    # This component FAILS on initial_env and PASSES on golden_env.
    # ----------------------------------------------------------------
    try:
        if ws_hr is None:
            print("FAIL: Component 2 — 'HR Data' sheet not found")
        else:
            originally_blank_correct = 0
            total_originally_blank = len(INITIAL_BLANK_ROWS_DEPT)
            mismatches = []

            for i, expected_dept in enumerate(EXPECTED_DEPARTMENT):
                row_idx = i + 2  # rows 2-21
                if row_idx not in INITIAL_BLANK_ROWS_DEPT:
                    continue  # Skip rows that were already filled (precondition)
                actual = ws_hr.cell(row=row_idx, column=2).value
                if actual is not None and str(actual).strip() == expected_dept:
                    originally_blank_correct += 1
                else:
                    mismatches.append(f"B{row_idx}: expected '{expected_dept}', got '{actual}'")

            if mismatches:
                for m in mismatches[:5]:
                    print(f"  FAIL: Department {m}")
                if len(mismatches) > 5:
                    print(f"  ... and {len(mismatches) - 5} more failures")

            if originally_blank_correct == total_originally_blank:
                print(f"PASS: Component 2 — All {total_originally_blank} originally-blank Department cells filled correctly (0.35 pts)")
                total_score += 0.35
            elif originally_blank_correct > 0:
                partial = round(0.35 * (originally_blank_correct / total_originally_blank), 4)
                print(f"PARTIAL: Component 2 — {originally_blank_correct}/{total_originally_blank} originally-blank Department cells filled correctly ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 2 — No originally-blank Department cells were filled correctly (0/{total_originally_blank})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ----------------------------------------------------------------
    # Component 3: Summary table with correct Division-Department-Employee Count (0.30 pts)
    # Task: create a summary table counting employees per Division-Department pair.
    # Initial state: only 1 sheet ('HR Data'), no Summary sheet.
    # Golden state: a new sheet exists with correct counts.
    # ----------------------------------------------------------------
    try:
        summary_sheet = None
        # Look for a sheet other than HR Data with Division/Department headers
        for sname in wb.sheetnames:
            if sname == 'HR Data':
                continue
            ws_candidate = wb[sname]
            h1 = ws_candidate.cell(row=1, column=1).value
            h2 = ws_candidate.cell(row=1, column=2).value
            h3 = ws_candidate.cell(row=1, column=3).value
            if (h1 and 'division' in str(h1).lower() and
                    h2 and 'department' in str(h2).lower() and
                    h3 is not None):
                summary_sheet = ws_candidate
                print(f"INFO: Found summary sheet: '{sname}'")
                break

        if summary_sheet is None:
            print(f"FAIL: Component 3 — No summary sheet with Division/Department headers found. Sheets: {wb.sheetnames}")
        else:
            # Read all rows from summary sheet (skip header)
            found_summary = {}
            for row in summary_sheet.iter_rows(min_row=2, max_row=summary_sheet.max_row, values_only=True):
                if len(row) < 3:
                    continue
                div_val, dept_val, count_val = row[0], row[1], row[2]
                if div_val is not None and dept_val is not None and count_val is not None:
                    key = (str(div_val).strip(), str(dept_val).strip())
                    try:
                        found_summary[key] = int(count_val)
                    except (ValueError, TypeError):
                        found_summary[key] = count_val

            # Compare with expected summary
            correct_pairs = 0
            total_pairs = len(EXPECTED_SUMMARY)
            for (div, dept), expected_count in EXPECTED_SUMMARY.items():
                key = (div, dept)
                if key in found_summary:
                    actual_count = found_summary[key]
                    if actual_count == expected_count:
                        correct_pairs += 1
                    else:
                        print(f"  FAIL: Summary ({div}, {dept}) — expected count {expected_count}, found {actual_count}")
                else:
                    print(f"  FAIL: Summary — missing pair ({div}, {dept})")

            if correct_pairs == total_pairs:
                print(f"PASS: Component 3 — All {total_pairs} Division-Department counts correct (0.30 pts)")
                total_score += 0.30
            elif correct_pairs > 0:
                partial = round(0.30 * (correct_pairs / total_pairs), 4)
                print(f"PARTIAL: Component 3 — {correct_pairs}/{total_pairs} pairs correct ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 3 — No Division-Department pairs matched in summary sheet")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in VM environment
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
