"""
Initial Setup: Consolidated P&L Statements from Three Business Units
Task ID: calc_fin_consolidated_pl_027
Domain: libreoffice_calc

Creates an Excel workbook with:
- BU_North, BU_South, BU_West sheets, each with 14 P&L line items and amounts
- Consolidated sheet with headers and line item names already filled (A2:A15)
  but NO formulas, NO color coding, NOT bold headers (those are the task)
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_consolidated_pl_027'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

# 14 standard P&L line items
LINE_ITEMS = [
    'Revenue',
    'COGS',
    'Gross Profit',
    'SG&A',
    'EBITDA',
    'D&A',
    'EBIT',
    'Interest Expense',
    'EBT',
    'Income Tax',
    'Net Income',
    'Other Income',
    'Operating Expenses',
    'Free Cash Flow',
]

# Realistic P&L amounts for each BU (in thousands USD)
BU_NORTH_DATA = [
    ('Revenue',           4_250_000),
    ('COGS',             -2_550_000),
    ('Gross Profit',      1_700_000),
    ('SG&A',              -510_000),
    ('EBITDA',            1_190_000),
    ('D&A',               -140_000),
    ('EBIT',              1_050_000),
    ('Interest Expense',   -85_000),
    ('EBT',                965_000),
    ('Income Tax',        -289_500),
    ('Net Income',         675_500),
    ('Other Income',        32_400),
    ('Operating Expenses', -645_000),
    ('Free Cash Flow',     820_000),
]

BU_SOUTH_DATA = [
    ('Revenue',           3_120_000),
    ('COGS',             -1_872_000),
    ('Gross Profit',      1_248_000),
    ('SG&A',              -374_400),
    ('EBITDA',             873_600),
    ('D&A',               -105_000),
    ('EBIT',               768_600),
    ('Interest Expense',   -62_000),
    ('EBT',                706_600),
    ('Income Tax',        -211_980),
    ('Net Income',         494_620),
    ('Other Income',        18_750),
    ('Operating Expenses', -479_000),
    ('Free Cash Flow',     601_000),
]

BU_WEST_DATA = [
    ('Revenue',           2_890_000),
    ('COGS',             -1_734_000),
    ('Gross Profit',      1_156_000),
    ('SG&A',              -346_800),
    ('EBITDA',             809_200),
    ('D&A',                -98_000),
    ('EBIT',               711_200),
    ('Interest Expense',   -57_500),
    ('EBT',                653_700),
    ('Income Tax',        -196_110),
    ('Net Income',         457_590),
    ('Other Income',        14_200),
    ('Operating Expenses', -441_000),
    ('Free Cash Flow',     556_500),
]


def create_bu_sheet(wb, sheet_name, data):
    """Create a Business Unit P&L sheet."""
    ws = wb.create_sheet(sheet_name)

    # Headers
    ws.cell(row=1, column=1, value='Line Item')
    ws.cell(row=1, column=2, value='Amount')

    # Style headers
    for col in range(1, 3):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_idx, (line_item, amount) in enumerate(data, 2):
        ws.cell(row=row_idx, column=1, value=line_item)
        ws.cell(row=row_idx, column=2, value=amount)
        ws.cell(row=row_idx, column=2).number_format = '$#,##0.00'

    # Column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16


def create_initial():
    wb = openpyxl.Workbook()

    # Remove default sheet
    default_ws = wb.active
    wb.remove(default_ws)

    # Create BU sheets
    create_bu_sheet(wb, 'BU_North', BU_NORTH_DATA)
    create_bu_sheet(wb, 'BU_South', BU_SOUTH_DATA)
    create_bu_sheet(wb, 'BU_West', BU_WEST_DATA)

    # Create Consolidated sheet
    ws_con = wb.create_sheet('Consolidated')

    # Row 1 headers (NOT bold in initial — bold is part of the task)
    headers = ['Line Item', 'BU North', 'BU South', 'BU West', 'Total',
               'North %', 'South %', 'West %']
    for col_idx, header in enumerate(headers, 1):
        cell = ws_con.cell(row=1, column=col_idx, value=header)
        cell.alignment = Alignment(horizontal='center')

    # A2:A15 line item names already entered (as stated in context)
    for row_idx, line_item in enumerate(LINE_ITEMS, 2):
        ws_con.cell(row=row_idx, column=1, value=line_item)

    # NOTE: B2:H15 are intentionally left EMPTY — formulas, colors, and bold
    # headers are the task the agent needs to complete.

    # Column widths
    ws_con.column_dimensions['A'].width = 22
    for col_letter in ['B', 'C', 'D', 'E']:
        ws_con.column_dimensions[col_letter].width = 16
    for col_letter in ['F', 'G', 'H']:
        ws_con.column_dimensions[col_letter].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheets: {wb.sheetnames}')
    print(f'Consolidated A2:A15 populated with {len(LINE_ITEMS)} line items')
    print('BU_North, BU_South, BU_West sheets populated with data')
    print('No formulas, no color coding, no bold headers in Consolidated (task elements)')


create_initial()
