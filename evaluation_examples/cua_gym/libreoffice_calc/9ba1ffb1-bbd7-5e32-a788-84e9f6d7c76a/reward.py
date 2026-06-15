"""
Reward Script: Consolidate P&L statements from three business units into a Consolidated sheet
Task ID: calc_fin_consolidated_pl_027
Domain: libreoffice_calc
Scoring:
  - Component 1: BU reference formulas in B2:B15, C2:C15, D2:D15 (0.30 pts)
  - Component 2: SUM total formulas in E2:E15 (0.20 pts)
  - Component 3: Percentage formulas (IFERROR) in F2:H15 with percentage number format (0.20 pts)
  - Component 4: Column color coding (B=light blue FFADD8E6, C=light green FF90EE90, D=light orange FFFFD580) (0.15 pts)
  - Component 5: Row 1 bold headers and currency format on B2:E15 (0.15 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_consolidated_pl_027'

# Expected fill colors for the three BU columns
BLUE_FILL   = 'FFADD8E6'  # light blue -> Column B (BU North)
GREEN_FILL  = 'FF90EE90'  # light green -> Column C (BU South)
ORANGE_FILL = 'FFFFD580'  # light orange -> Column D (BU West)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Consolidated sheet must exist
    if 'Consolidated' not in wb.sheetnames:
        print("FAIL: 'Consolidated' sheet not found in workbook")
        print("Score: 0.0/1.0")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Consolidated']

    # -----------------------------------------------------------------------
    # Component 1: BU reference formulas in B2:B15, C2:C15, D2:D15 (0.30 pts)
    # B2:B15 = =BU_North!Bx, C2:C15 = =BU_South!Bx, D2:D15 = =BU_West!Bx
    # We count how many of the 14*3=42 cells have the correct cross-sheet formula.
    # -----------------------------------------------------------------------
    try:
        ref_ok = 0
        ref_total = 42  # 14 rows x 3 columns

        for row in range(2, 16):
            # Column B: must reference BU_North column B
            b_val = ws.cell(row=row, column=2).value
            if isinstance(b_val, str) and 'BU_North' in b_val and f'B{row}' in b_val:
                ref_ok += 1

            # Column C: must reference BU_South column B
            c_val = ws.cell(row=row, column=3).value
            if isinstance(c_val, str) and 'BU_South' in c_val and f'B{row}' in c_val:
                ref_ok += 1

            # Column D: must reference BU_West column B
            d_val = ws.cell(row=row, column=4).value
            if isinstance(d_val, str) and 'BU_West' in d_val and f'B{row}' in d_val:
                ref_ok += 1

        comp1_partial = round(ref_ok / ref_total * 0.30, 4)
        if ref_ok == ref_total:
            print(f"PASS: Component 1 — All {ref_total} BU reference formulas present (0.30 pts)")
            total_score += 0.30
        elif ref_ok > 0:
            print(f"FAIL: Component 1 — BU reference formulas incomplete ({ref_ok}/{ref_total}); partial={comp1_partial}")
            total_score += comp1_partial
        else:
            print(f"FAIL: Component 1 — No BU reference formulas found (0/{ref_total})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: SUM total formulas in E2:E15 (0.20 pts)
    # E2:E15 = =SUM(Bx:Dx) — sum of the three BU columns
    # -----------------------------------------------------------------------
    try:
        e_ok = 0
        for row in range(2, 16):
            e_val = ws.cell(row=row, column=5).value
            if isinstance(e_val, str):
                normalized = e_val.upper().replace(' ', '')
                if normalized.startswith('=SUM(') and f'B{row}' in e_val and f'D{row}' in e_val:
                    e_ok += 1

        comp2_partial = round(e_ok / 14 * 0.20, 4)
        if e_ok == 14:
            print(f"PASS: Component 2 — All 14 SUM total formulas present in E2:E15 (0.20 pts)")
            total_score += 0.20
        elif e_ok > 0:
            print(f"FAIL: Component 2 — SUM formulas incomplete ({e_ok}/14); partial={comp2_partial}")
            total_score += comp2_partial
        else:
            print(f"FAIL: Component 2 — No SUM total formulas found in column E")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Percentage formulas with IFERROR in F2:H15, formatted as % (0.20 pts)
    # F2:F15 = =IFERROR(Bx/Ex,0), G2:G15 = =IFERROR(Cx/Ex,0), H2:H15 = =IFERROR(Dx/Ex,0)
    # All formatted with a percentage number format (e.g. '0.00%')
    # -----------------------------------------------------------------------
    try:
        fgh_formula_ok = 0
        fgh_format_ok = 0
        fgh_total = 42  # 14 rows x 3 columns (F, G, H)

        for row in range(2, 16):
            # Column F: IFERROR(Bx/Ex, ...)
            f_val = ws.cell(row=row, column=6).value
            if isinstance(f_val, str) and 'IFERROR' in f_val.upper() and f'B{row}' in f_val and f'E{row}' in f_val:
                fgh_formula_ok += 1
            if '%' in str(ws.cell(row=row, column=6).number_format):
                fgh_format_ok += 1

            # Column G: IFERROR(Cx/Ex, ...)
            g_val = ws.cell(row=row, column=7).value
            if isinstance(g_val, str) and 'IFERROR' in g_val.upper() and f'C{row}' in g_val and f'E{row}' in g_val:
                fgh_formula_ok += 1
            if '%' in str(ws.cell(row=row, column=7).number_format):
                fgh_format_ok += 1

            # Column H: IFERROR(Dx/Ex, ...)
            h_val = ws.cell(row=row, column=8).value
            if isinstance(h_val, str) and 'IFERROR' in h_val.upper() and f'D{row}' in h_val and f'E{row}' in h_val:
                fgh_formula_ok += 1
            if '%' in str(ws.cell(row=row, column=8).number_format):
                fgh_format_ok += 1

        # Score: 0.14 for formulas + 0.06 for formats
        formula_ratio = fgh_formula_ok / fgh_total
        format_ratio = fgh_format_ok / fgh_total
        comp3_partial = round(formula_ratio * 0.14 + format_ratio * 0.06, 4)

        if fgh_formula_ok == fgh_total and fgh_format_ok == fgh_total:
            print(f"PASS: Component 3 — All {fgh_total} IFERROR percentage formulas present and formatted (0.20 pts)")
            total_score += 0.20
        elif fgh_formula_ok > 0 or fgh_format_ok > 0:
            print(f"FAIL: Component 3 — IFERROR formulas ({fgh_formula_ok}/{fgh_total}) or formats ({fgh_format_ok}/{fgh_total}) incomplete; partial={comp3_partial}")
            total_score += comp3_partial
        else:
            print(f"FAIL: Component 3 — No IFERROR percentage formulas found in columns F/G/H")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: Column color coding (0.15 pts)
    # Col B (BU North) = light blue  FFADD8E6
    # Col C (BU South) = light green FF90EE90
    # Col D (BU West)  = light orange FFFFD580
    # Verify data rows 2-15 (header row excluded as it may not be colored)
    # -----------------------------------------------------------------------
    try:
        color_ok = 0
        color_total = 42  # 14 rows x 3 columns

        for row in range(2, 16):
            try:
                if ws.cell(row=row, column=2).fill.fgColor.rgb == BLUE_FILL:
                    color_ok += 1
            except Exception:
                pass

            try:
                if ws.cell(row=row, column=3).fill.fgColor.rgb == GREEN_FILL:
                    color_ok += 1
            except Exception:
                pass

            try:
                if ws.cell(row=row, column=4).fill.fgColor.rgb == ORANGE_FILL:
                    color_ok += 1
            except Exception:
                pass

        comp4_partial = round(color_ok / color_total * 0.15, 4)
        if color_ok == color_total:
            print(f"PASS: Component 4 — All {color_total} color-coded cells correct (B=blue, C=green, D=orange) (0.15 pts)")
            total_score += 0.15
        elif color_ok > 0:
            print(f"FAIL: Component 4 — Column colors incomplete ({color_ok}/{color_total}); partial={comp4_partial}")
            total_score += comp4_partial
        else:
            print(f"FAIL: Component 4 — No column color coding found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -----------------------------------------------------------------------
    # Component 5: Row 1 bold headers AND currency format on B2:E15 (0.15 pts)
    # All 8 header cells (A1:H1) must be bold.
    # B2:E15 data cells must use currency format ($#,##0.00 or similar).
    # -----------------------------------------------------------------------
    try:
        bold_ok = 0
        for col in range(1, 9):
            if ws.cell(row=1, column=col).font.bold:
                bold_ok += 1

        currency_ok = 0
        currency_total = 0
        for row in range(2, 16):
            for col in range(2, 6):  # columns B (2) through E (5)
                currency_total += 1
                fmt = ws.cell(row=row, column=col).number_format
                if '$' in str(fmt) or ('#,##0' in str(fmt)):
                    currency_ok += 1

        bold_ratio = bold_ok / 8
        currency_ratio = currency_ok / currency_total
        comp5_partial = round(bold_ratio * 0.08 + currency_ratio * 0.07, 4)

        if bold_ok == 8 and currency_ok == currency_total:
            print(f"PASS: Component 5 — Row 1 bold ({bold_ok}/8) and currency format ({currency_ok}/{currency_total} cells) (0.15 pts)")
            total_score += 0.15
        elif bold_ok > 0 or currency_ok > 0:
            print(f"FAIL: Component 5 — Row 1 bold ({bold_ok}/8) or currency format ({currency_ok}/{currency_total} cells) incomplete; partial={comp5_partial}")
            total_score += comp5_partial
        else:
            print(f"FAIL: Component 5 — Neither bold headers nor currency format found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
