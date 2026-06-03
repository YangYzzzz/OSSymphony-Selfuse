"""
Reward Script: Set up grade report for printing
Task ID: calc_edu_print_grade_report_010
Domain: libreoffice_calc
Scoring:
  - Component 1: Print area set to A1:Z63 (0.25 pts)
  - Component 2: Page orientation is landscape (0.20 pts)
  - Component 3: Print title rows set to $3:$3 (0.20 pts)
  - Component 4: Print title columns set to $A:$A (0.15 pts)
  - Component 5: Page header center text is 'MATH 101 Grade Report - Fall 2025' (0.20 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_edu_print_grade_report_010'
SHEET_NAME = 'Grade Report'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the sheet exists
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Component 1: Print area set to A1:Z63 (0.25 points)
    # Initial state: no print area set (empty string)
    # Expected golden: 'Grade Report'!$A$1:$Z$63
    try:
        print_area = ws.print_area
        # Accept both formats: with/without sheet name prefix, with/without $ signs
        normalized = str(print_area).replace("'Grade Report'!", "").replace("$", "").upper() if print_area else ""
        if normalized and ("A1:Z63" in normalized or normalized == "A1:Z63"):
            print(f"PASS: Component 1 — Print area is set to A1:Z63 (found: {print_area}) (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — Expected print area A1:Z63, found: {repr(print_area)}")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not check print area: {e}")

    # Component 2: Page orientation is landscape (0.20 points)
    # Initial state: orientation is None (default/portrait)
    # Expected golden: 'landscape'
    try:
        orientation = ws.page_setup.orientation
        if orientation and str(orientation).lower() == 'landscape':
            print(f"PASS: Component 2 — Page orientation is landscape (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 — Expected orientation 'landscape', found: {repr(orientation)}")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not check page orientation: {e}")

    # Component 3: Print title rows set to $3:$3 (0.20 points)
    # Initial state: print_title_rows is None
    # Expected golden: '$3:$3' (row 3 repeats at the top of each printed page)
    try:
        title_rows = ws.print_title_rows
        # Normalize: remove $ signs and compare
        normalized_rows = str(title_rows).replace("$", "") if title_rows else ""
        if title_rows and ("3:3" in normalized_rows):
            print(f"PASS: Component 3 — Print title rows set to $3:$3 (found: {title_rows}) (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 3 — Expected print title rows '$3:$3', found: {repr(title_rows)}")
    except Exception as e:
        print(f"ERROR: Component 3 — Could not check print title rows: {e}")

    # Component 4: Print title columns set to $A:$A (0.15 points)
    # Initial state: print_title_cols is None
    # Expected golden: '$A:$A' (column A repeats on the left of each printed page)
    try:
        title_cols = ws.print_title_cols
        # Normalize: remove $ signs and compare
        normalized_cols = str(title_cols).replace("$", "") if title_cols else ""
        if title_cols and ("A:A" in normalized_cols):
            print(f"PASS: Component 4 — Print title columns set to $A:$A (found: {title_cols}) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — Expected print title cols '$A:$A', found: {repr(title_cols)}")
    except Exception as e:
        print(f"ERROR: Component 4 — Could not check print title columns: {e}")

    # Component 5: Page header center text is 'MATH 101 Grade Report - Fall 2025' (0.20 points)
    # Initial state: oddHeader.center.text is None
    # Expected golden: 'MATH 101 Grade Report - Fall 2025'
    try:
        header_center = ws.oddHeader.center.text
        expected_header = 'MATH 101 Grade Report - Fall 2025'
        if header_center and str(header_center).strip() == expected_header:
            print(f"PASS: Component 5 — Page header center text matches '{expected_header}' (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 5 — Expected header center text '{expected_header}', found: {repr(header_center)}")
    except Exception as e:
        print(f"ERROR: Component 5 — Could not check page header: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
