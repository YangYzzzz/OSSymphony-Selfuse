"""
Initial Setup: Grade report spreadsheet for MATH 101 - Algebra Fundamentals
Task ID: calc_edu_print_grade_report_010
Domain: libreoffice_calc

Creates a grade report spreadsheet with:
- Sheet 'Grade Report' with class title, semester, header row, 60 student rows
- Assignment columns B through Z (25 columns)
- Scratch columns AA-AE
- NO print area, headers, orientation, or row/column titles set
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_print_grade_report_010'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Grade Report'

    # Row 1: Class title
    ws['A1'] = 'MATH 101 - Algebra Fundamentals'
    ws['A1'].font = Font(bold=True, size=14)

    # Row 2: Semester label
    ws['A2'] = 'Fall 2025'
    ws['A2'].font = Font(italic=True, size=11)

    # Row 3: Header row — Student Name + 25 assignment columns (B-Z)
    assignment_headers = (
        ['Student Name'] +
        [f'HW{i}' for i in range(1, 8)] +   # HW1-HW7 (7 cols)
        [f'Quiz{i}' for i in range(1, 6)] +  # Quiz1-Quiz5 (5 cols)
        [f'Exam{i}' for i in range(1, 4)] +  # Exam1-Exam3 (3 cols)
        ['Midterm', 'Final'] +               # 2 cols
        ['Lab1', 'Lab2', 'Lab3'] +           # 3 cols
        ['Project1', 'Project2'] +           # 2 cols
        ['Participation', 'Extra Credit']    # 2 cols
        # Total: 1 + 7 + 5 + 3 + 2 + 3 + 2 + 2 + 1 = 26 cols (A-Z)
    )
    # That's exactly 26 columns (A=Student Name, B-Z=25 assignment columns)
    for col_idx, header in enumerate(assignment_headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    ws.column_dimensions['A'].width = 22

    # Realistic student data (60 students, rows 4-63)
    first_names = [
        'Emma', 'Liam', 'Olivia', 'Noah', 'Ava', 'William', 'Sophia', 'James',
        'Isabella', 'Oliver', 'Mia', 'Benjamin', 'Charlotte', 'Elijah', 'Amelia',
        'Lucas', 'Harper', 'Mason', 'Evelyn', 'Logan', 'Abigail', 'Alexander',
        'Emily', 'Ethan', 'Elizabeth', 'Jacob', 'Sofia', 'Michael', 'Avery',
        'Daniel', 'Ella', 'Henry', 'Scarlett', 'Jackson', 'Grace', 'Sebastian',
        'Chloe', 'Aiden', 'Victoria', 'Matthew', 'Riley', 'Samuel', 'Aria',
        'David', 'Lily', 'Joseph', 'Zoey', 'Carter', 'Penelope', 'Owen',
        'Layla', 'Wyatt', 'Nora', 'John', 'Lillian', 'Jack', 'Eleanor',
        'Luke', 'Hannah', 'Jayden', 'Lilliana'
    ]

    last_names = [
        'Johnson', 'Williams', 'Brown', 'Jones', 'Garcia', 'Miller', 'Davis',
        'Martinez', 'Hernandez', 'Lopez', 'Gonzalez', 'Wilson', 'Anderson',
        'Thomas', 'Taylor', 'Moore', 'Jackson', 'Martin', 'Lee', 'Perez',
        'Thompson', 'White', 'Harris', 'Sanchez', 'Clark', 'Ramirez', 'Lewis',
        'Robinson', 'Walker', 'Young', 'Allen', 'King', 'Wright', 'Scott',
        'Torres', 'Nguyen', 'Hill', 'Flores', 'Green', 'Adams', 'Nelson',
        'Baker', 'Hall', 'Rivera', 'Campbell', 'Mitchell', 'Carter', 'Roberts',
        'Gomez', 'Phillips', 'Evans', 'Turner', 'Diaz', 'Parker', 'Cruz',
        'Edwards', 'Collins', 'Reyes', 'Stewart', 'Morris', 'Morales'
    ]

    import random
    random.seed(42)

    for i in range(60):
        row = i + 4
        student_name = f'{first_names[i]} {last_names[i]}'
        ws.cell(row=row, column=1, value=student_name)

        # HW1-HW7: out of 10
        for col in range(2, 9):
            ws.cell(row=row, column=col, value=round(random.uniform(6.0, 10.0), 1))

        # Quiz1-Quiz5: out of 20
        for col in range(9, 14):
            ws.cell(row=row, column=col, value=round(random.uniform(12.0, 20.0), 1))

        # Exam1-Exam3: out of 100
        for col in range(14, 17):
            ws.cell(row=row, column=col, value=round(random.uniform(60.0, 100.0), 1))

        # Midterm: out of 100
        ws.cell(row=row, column=17, value=round(random.uniform(65.0, 98.0), 1))

        # Final: out of 100
        ws.cell(row=row, column=18, value=round(random.uniform(60.0, 100.0), 1))

        # Lab1-Lab3: out of 25
        for col in range(19, 22):
            ws.cell(row=row, column=col, value=round(random.uniform(18.0, 25.0), 1))

        # Project1-Project2: out of 50
        for col in range(22, 24):
            ws.cell(row=row, column=col, value=round(random.uniform(35.0, 50.0), 1))

        # Participation: out of 10
        ws.cell(row=row, column=24, value=round(random.uniform(6.0, 10.0), 1))

        # Extra Credit: out of 5
        ws.cell(row=row, column=25, value=round(random.uniform(0.0, 5.0), 1))

        # Column Z: calculated total (just raw numeric, not a formula)
        ws.cell(row=row, column=26, value=round(random.uniform(75.0, 98.0), 1))

    # Header for column Z
    ws.cell(row=3, column=26, value='Total %')
    ws.cell(row=3, column=26).font = Font(bold=True)
    ws.cell(row=3, column=26).fill = PatternFill(
        start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    ws.cell(row=3, column=26).alignment = Alignment(horizontal='center')

    # Columns AA-AE: scratch work calculations (cols 27-31)
    scratch_headers = ['Weighted HW', 'Weighted Quiz', 'Weighted Exam', 'Curve Points', 'Adj. Total']
    for col_idx, hdr in enumerate(scratch_headers, 27):
        cell = ws.cell(row=3, column=col_idx, value=hdr)
        cell.font = Font(italic=True, color='FF808080')
    for i in range(60):
        row = i + 4
        for col in range(27, 32):
            ws.cell(row=row, column=col, value=round(random.uniform(50.0, 95.0), 2))

    # NOTE: NO print area set, NO page headers, NO orientation, NO row/column titles
    # These are all task-completion items that must NOT be in initial state

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheets: {wb.sheetnames}')
    print(f'  Grade Report: rows 1-63, columns A-AE')
    print(f'  Print settings: NONE set (as required)')


create_initial()
