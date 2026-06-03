"""
Initial Setup: VAT Calculation Worksheet for European Sales
Task ID: calc_fin_vat_calculation_055
Domain: libreoffice_calc

Creates the pre-task state:
- Sheet 'VATCalc' with headers in row 1
- Rows 2-60: Invoice#, Country, Net Amount filled; VAT Rate, VAT Amount, Gross Amount EMPTY
- H1:I6 VAT rate reference table
- Row 1 bold and frozen (already set as initial state)
- NO formulas in D/E/F columns
- NO data validation on B column
- NO summary section
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_vat_calculation_055'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'VATCalc'

    # --- Row 1: Headers ---
    headers = ['Invoice#', 'Country', 'Net Amount', 'VAT Rate', 'VAT Amount', 'Gross Amount']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Freeze row 1
    ws.freeze_panes = 'A2'

    # --- VAT reference table in H:I (columns 8 and 9) ---
    ws.cell(row=1, column=8, value='Country').font = Font(bold=True)
    ws.cell(row=1, column=9, value='VAT Rate').font = Font(bold=True)

    vat_rates = [
        ('Germany', 0.19),
        ('France', 0.20),
        ('UK', 0.20),
        ('Netherlands', 0.21),
        ('Spain', 0.21),
    ]
    for i, (country, rate) in enumerate(vat_rates, 2):
        ws.cell(row=i, column=8, value=country)
        rate_cell = ws.cell(row=i, column=9, value=rate)
        rate_cell.number_format = '0%'

    # --- Invoice data: rows 2-60 ---
    countries = ['Germany', 'France', 'UK', 'Netherlands', 'Spain']
    # Realistic company/product invoice data
    invoice_data = [
        ('INV-2025-001', 'Germany', 1250.00),
        ('INV-2025-002', 'France', 3480.50),
        ('INV-2025-003', 'UK', 780.00),
        ('INV-2025-004', 'Netherlands', 5200.75),
        ('INV-2025-005', 'Spain', 920.00),
        ('INV-2025-006', 'Germany', 2100.00),
        ('INV-2025-007', 'France', 640.25),
        ('INV-2025-008', 'UK', 1890.00),
        ('INV-2025-009', 'Netherlands', 3350.00),
        ('INV-2025-010', 'Spain', 445.50),
        ('INV-2025-011', 'Germany', 7800.00),
        ('INV-2025-012', 'France', 1100.00),
        ('INV-2025-013', 'UK', 2950.75),
        ('INV-2025-014', 'Netherlands', 660.00),
        ('INV-2025-015', 'Spain', 3700.00),
        ('INV-2025-016', 'Germany', 530.00),
        ('INV-2025-017', 'France', 4200.50),
        ('INV-2025-018', 'UK', 1500.00),
        ('INV-2025-019', 'Netherlands', 2750.00),
        ('INV-2025-020', 'Spain', 890.25),
        ('INV-2025-021', 'Germany', 3100.00),
        ('INV-2025-022', 'France', 720.00),
        ('INV-2025-023', 'UK', 4450.00),
        ('INV-2025-024', 'Netherlands', 1350.75),
        ('INV-2025-025', 'Spain', 6200.00),
        ('INV-2025-026', 'Germany', 980.50),
        ('INV-2025-027', 'France', 2300.00),
        ('INV-2025-028', 'UK', 500.00),
        ('INV-2025-029', 'Netherlands', 4100.00),
        ('INV-2025-030', 'Spain', 1750.00),
        ('INV-2025-031', 'Germany', 8500.00),
        ('INV-2025-032', 'France', 340.00),
        ('INV-2025-033', 'UK', 2650.50),
        ('INV-2025-034', 'Netherlands', 1900.00),
        ('INV-2025-035', 'Spain', 430.00),
        ('INV-2025-036', 'Germany', 1650.00),
        ('INV-2025-037', 'France', 5300.75),
        ('INV-2025-038', 'UK', 760.00),
        ('INV-2025-039', 'Netherlands', 2200.00),
        ('INV-2025-040', 'Spain', 3850.00),
        ('INV-2025-041', 'Germany', 2450.50),
        ('INV-2025-042', 'France', 1080.00),
        ('INV-2025-043', 'UK', 6700.00),
        ('INV-2025-044', 'Netherlands', 510.75),
        ('INV-2025-045', 'Spain', 1200.00),
        ('INV-2025-046', 'Germany', 3900.00),
        ('INV-2025-047', 'France', 2800.50),
        ('INV-2025-048', 'UK', 940.00),
        ('INV-2025-049', 'Netherlands', 7200.00),
        ('INV-2025-050', 'Spain', 1560.25),
        ('INV-2025-051', 'Germany', 620.00),
        ('INV-2025-052', 'France', 4750.00),
        ('INV-2025-053', 'UK', 1430.50),
        ('INV-2025-054', 'Netherlands', 3000.00),
        ('INV-2025-055', 'Spain', 2100.00),
        ('INV-2025-056', 'Germany', 5500.75),
        ('INV-2025-057', 'France', 870.00),
        ('INV-2025-058', 'UK', 1950.00),
        ('INV-2025-059', 'Netherlands', 480.50),
    ]

    for r, (inv_num, country, net_amount) in enumerate(invoice_data, 2):
        ws.cell(row=r, column=1, value=inv_num)
        ws.cell(row=r, column=2, value=country)
        net_cell = ws.cell(row=r, column=3, value=net_amount)
        net_cell.number_format = '#,##0.00'
        # Columns D (VAT Rate), E (VAT Amount), F (Gross Amount) are intentionally left EMPTY

    # Set column widths for readability
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets:', wb.sheetnames)
    print('Row 1 headers: Invoice#, Country, Net Amount, VAT Rate, VAT Amount, Gross Amount')
    print('Data rows: 2-60 (A, B, C filled; D, E, F empty)')
    print('VAT reference table: H1:I6')

create_initial()
