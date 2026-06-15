"""
Reward Script: VAT Calculation Worksheet Setup
Task ID: calc_fin_vat_calculation_055
Domain: libreoffice_calc
Scoring:
  - Component 1: VLOOKUP formulas in D2:D60 with percentage format (0.25 pts)
  - Component 2: VAT amount formulas =C*D in E2:E60 with currency format (0.20 pts)
  - Component 3: Gross amount formulas =C+E in F2:F60 with currency format (0.20 pts)
  - Component 4: Data validation dropdown on B2:B60 sourced from $H$2:$H$6 (0.15 pts)
  - Component 5: Summary section at rows 62-67 with SUMIF formulas, bold, currency (0.20 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_vat_calculation_055'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the VATCalc sheet exists
    if 'VATCalc' not in wb.sheetnames:
        print("CRITICAL: Sheet 'VATCalc' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['VATCalc']

    # -------------------------------------------------------------------------
    # Component 1: VLOOKUP formulas in D2:D60 with percentage format (0.25 pts)
    # FAILS on initial (D column is empty) → PASSES on golden (VLOOKUP formulas present)
    # -------------------------------------------------------------------------
    try:
        vlookup_count = 0
        pct_format_count = 0
        for row in range(2, 61):  # rows 2 to 60
            cell_d = ws.cell(row=row, column=4)
            val = cell_d.value
            if val and isinstance(val, str) and 'VLOOKUP' in val.upper():
                vlookup_count += 1
            # Check percentage format - should be '0%' or similar
            fmt = cell_d.number_format
            if fmt and ('%' in fmt or 'percent' in fmt.lower()):
                pct_format_count += 1

        # Require at least 55 of 59 rows have VLOOKUP formula
        if vlookup_count >= 55:
            print(f"PASS: Component 1a — VLOOKUP formulas found in D column ({vlookup_count}/59 rows)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1a — Expected VLOOKUP formulas in D2:D60, found in {vlookup_count}/59 rows")

        if pct_format_count >= 55:
            print(f"PASS: Component 1b — Percentage format in D column ({pct_format_count}/59 rows)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1b — Expected percentage format in D2:D60, found in {pct_format_count}/59 rows")

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: VAT amount formulas in E2:E60 with currency format (0.20 pts)
    # Formula should be =C*D; currency format #,##0.00 or similar
    # FAILS on initial (E column is empty) → PASSES on golden
    # -------------------------------------------------------------------------
    try:
        vat_formula_count = 0
        vat_currency_count = 0
        for row in range(2, 61):
            cell_e = ws.cell(row=row, column=5)
            val = cell_e.value
            # Formula should multiply C and D (e.g., =C2*D2)
            if val and isinstance(val, str):
                val_upper = val.upper().replace(' ', '')
                if '*' in val_upper and 'C' in val_upper and 'D' in val_upper:
                    vat_formula_count += 1
            # Check currency format
            fmt = cell_e.number_format
            if fmt and ('#,##0' in fmt or '$' in fmt or '€' in fmt or '£' in fmt):
                vat_currency_count += 1

        if vat_formula_count >= 55:
            print(f"PASS: Component 2a — VAT amount formulas in E column ({vat_formula_count}/59 rows)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2a — Expected C*D formulas in E2:E60, found in {vat_formula_count}/59 rows")

        if vat_currency_count >= 55:
            print(f"PASS: Component 2b — Currency format in E column ({vat_currency_count}/59 rows)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2b — Expected currency format in E2:E60, found in {vat_currency_count}/59 rows")

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Gross amount formulas in F2:F60 with currency format (0.20 pts)
    # Formula should add C and E (e.g., =C2+E2)
    # FAILS on initial (F column is empty) → PASSES on golden
    # -------------------------------------------------------------------------
    try:
        gross_formula_count = 0
        gross_currency_count = 0
        for row in range(2, 61):
            cell_f = ws.cell(row=row, column=6)
            val = cell_f.value
            if val and isinstance(val, str):
                val_upper = val.upper().replace(' ', '')
                if '+' in val_upper and 'C' in val_upper and 'E' in val_upper:
                    gross_formula_count += 1
            fmt = cell_f.number_format
            if fmt and ('#,##0' in fmt or '$' in fmt or '€' in fmt or '£' in fmt):
                gross_currency_count += 1

        if gross_formula_count >= 55:
            print(f"PASS: Component 3a — Gross amount formulas in F column ({gross_formula_count}/59 rows)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3a — Expected C+E formulas in F2:F60, found in {gross_formula_count}/59 rows")

        if gross_currency_count >= 55:
            print(f"PASS: Component 3b — Currency format in F column ({gross_currency_count}/59 rows)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3b — Expected currency format in F2:F60, found in {gross_currency_count}/59 rows")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: Data validation dropdown on B2:B60 sourced from $H$2:$H$6 (0.15 pts)
    # FAILS on initial (no data validations) → PASSES on golden
    # -------------------------------------------------------------------------
    try:
        dvs = ws.data_validations.dataValidation
        dv_found = False
        dv_correct_formula = False
        dv_covers_b_col = False

        for dv in dvs:
            if dv.type == 'list':
                dv_found = True
                # Check formula references H2:H6
                if dv.formula1 and 'H' in dv.formula1.upper() and ('2' in dv.formula1):
                    dv_correct_formula = True
                # Check it covers column B rows 2-60
                sqref_str = str(dv.sqref)
                if 'B' in sqref_str.upper():
                    dv_covers_b_col = True

        if dv_found and dv_correct_formula and dv_covers_b_col:
            print(f"PASS: Component 4 — Data validation dropdown on B2:B60 from $H$2:$H$6")
            total_score += 0.15
        elif dv_found:
            print(f"FAIL: Component 4 — Data validation found but incorrect formula ({dvs[0].formula1}) or range ({dvs[0].sqref})")
        else:
            print(f"FAIL: Component 4 — No list data validation found on B column")

    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -------------------------------------------------------------------------
    # Component 5: Summary section at rows 62-67 (0.20 pts)
    # A62='VAT Summary by Country', A63:A67=country names, B63:B67=SUMIF formulas
    # bold+currency format on B63:B67
    # FAILS on initial (rows 62+ are empty) → PASSES on golden
    # -------------------------------------------------------------------------
    try:
        a62_val = ws.cell(row=62, column=1).value
        header_ok = a62_val is not None and 'VAT' in str(a62_val).upper()

        country_names = ['Germany', 'France', 'UK', 'Netherlands', 'Spain']
        countries_found = 0
        sumif_count = 0
        bold_currency_count = 0

        for idx, row in enumerate(range(63, 68)):
            a_cell = ws.cell(row=row, column=1)
            b_cell = ws.cell(row=row, column=2)
            a_val = a_cell.value
            b_val = b_cell.value

            # Check country name in A column
            if a_val and any(c.lower() in str(a_val).lower() for c in country_names):
                countries_found += 1

            # Check SUMIF formula in B column
            if b_val and isinstance(b_val, str) and 'SUMIF' in b_val.upper():
                sumif_count += 1

            # Check bold and currency format on B column
            b_bold = b_cell.font.bold
            b_fmt = b_cell.number_format
            if b_bold and b_fmt and '#,##0' in b_fmt:
                bold_currency_count += 1

        if header_ok:
            print(f"PASS: Component 5a — VAT summary header in A62 ('{a62_val}')")
        else:
            print(f"FAIL: Component 5a — Expected 'VAT Summary by Country' in A62, found: {repr(a62_val)}")

        if countries_found >= 4 and sumif_count >= 4:
            print(f"PASS: Component 5b — Country names ({countries_found}/5) and SUMIF formulas ({sumif_count}/5) in rows 63-67")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5b — Countries found: {countries_found}/5, SUMIF formulas: {sumif_count}/5")

        if bold_currency_count >= 4:
            print(f"PASS: Component 5c — Bold+currency format on B63:B67 ({bold_currency_count}/5)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5c — Expected bold+currency on B63:B67, found: {bold_currency_count}/5 rows")

    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
