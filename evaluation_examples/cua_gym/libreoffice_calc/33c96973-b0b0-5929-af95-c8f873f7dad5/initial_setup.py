"""
Initial Setup: Design a time-off request calendar with employee data across a monthly grid.
Task ID: calc_hr_064
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_064'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calendar"

    # Row 1: Headers - A1='Employee', B1-AF1: dates 1-31
    ws.cell(row=1, column=1, value="Employee")
    for day in range(1, 32):
        ws.cell(row=1, column=day + 1, value=day)

    # Employee names in A2:A5
    employees = ['Alice', 'Bob', 'Carol', 'Dan']
    for i, name in enumerate(employees, 2):
        ws.cell(row=i, column=1, value=name)

    # Time-off data: realistic mix of A, P, D, and blank
    # Alice: has a week approved vacation, some pending days, one denied
    alice_data = {
        2: 'A', 3: 'A', 4: 'A', 5: 'A', 6: 'A',   # Mon-Fri week 1 vacation
        10: 'P',                                        # pending day
        15: 'D',                                        # denied request
        22: 'A', 23: 'A',                              # two approved days
        28: 'P',                                        # pending
    }

    # Bob: scattered approved and pending days
    bob_data = {
        1: 'A',                                         # approved day off
        7: 'P', 8: 'P',                                # pending two days
        14: 'A',                                        # approved
        19: 'D', 20: 'D',                              # denied two days
        25: 'A', 26: 'A', 27: 'A',                    # approved 3 days
    }

    # Carol: mostly pending requests
    carol_data = {
        3: 'P', 4: 'P', 5: 'P',                       # pending 3 days
        11: 'A',                                        # approved
        16: 'D',                                        # denied
        21: 'P', 22: 'P',                              # pending
        29: 'A', 30: 'A', 31: 'A',                    # approved end of month
    }

    # Dan: mix of everything
    dan_data = {
        2: 'D',                                         # denied
        6: 'A', 7: 'A',                                # approved weekend adjacent
        12: 'P',                                        # pending
        17: 'A', 18: 'A', 19: 'A',                    # approved mid-month
        24: 'D',                                        # denied
        30: 'P', 31: 'P',                              # pending end of month
    }

    all_data = [alice_data, bob_data, carol_data, dan_data]
    for row_idx, emp_data in enumerate(all_data, 2):
        for day, status in emp_data.items():
            ws.cell(row=row_idx, column=day + 1, value=status)

    # Style headers for readability
    from openpyxl.styles import Font, Alignment
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center")

    for col in range(1, 33):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = center_align

    # Set column widths
    ws.column_dimensions['A'].width = 12
    for day in range(1, 32):
        col_letter = openpyxl.utils.get_column_letter(day + 1)
        ws.column_dimensions[col_letter].width = 4

    # Center-align data cells
    for row in range(2, 6):
        for col in range(2, 33):
            ws.cell(row=row, column=col).alignment = center_align

    # NO conditional formatting in initial - that's the task!

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
