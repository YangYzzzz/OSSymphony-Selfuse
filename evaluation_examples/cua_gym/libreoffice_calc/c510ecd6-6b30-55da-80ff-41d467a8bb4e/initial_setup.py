"""
Initial Setup: Highlight weekend rows in patient appointment schedule
Task ID: osworld_calc_conditional_format_weekday_008
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_conditional_format_weekday_008'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Appointments ---
    ws = wb.active
    ws.title = 'Appointments'

    # Headers
    headers = ['Appointment Date', 'Patient ID', 'Doctor', 'Appointment Type', 'Duration (min)']
    header_font = Font(name='Calibri', bold=True, size=11)
    header_fill = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[1].height = 22

    # Appointment data — realistic mix of weekdays and weekends
    # Dates chosen to include Saturdays (weekday=7 in LibreOffice WEEKDAY) and
    # Sundays (weekday=1 in LibreOffice WEEKDAY default mode)
    appointment_data = [
        # (date,        patient_id, doctor,               appt_type,          duration)
        (date(2025, 3, 3),  'P-10421', 'Dr. Sarah Mitchell',  'General Checkup',    30),   # Monday
        (date(2025, 3, 4),  'P-10835', 'Dr. James Okonkwo',   'Follow-up',          20),   # Tuesday
        (date(2025, 3, 5),  'P-11204', 'Dr. Elena Vasquez',   'Blood Work Review',  45),   # Wednesday
        (date(2025, 3, 6),  'P-10567', 'Dr. Sarah Mitchell',  'Cardiology Consult', 60),   # Thursday
        (date(2025, 3, 7),  'P-11089', 'Dr. James Okonkwo',   'Physical Therapy',   50),   # Friday
        (date(2025, 3, 8),  'P-10342', 'Dr. Priya Nair',      'Urgent Care',        30),   # Saturday (weekend)
        (date(2025, 3, 9),  'P-11567', 'Dr. Priya Nair',      'Emergency Consult',  45),   # Sunday (weekend)
        (date(2025, 3, 10), 'P-10923', 'Dr. Elena Vasquez',   'Post-Op Review',     40),   # Monday
        (date(2025, 3, 11), 'P-11312', 'Dr. Sarah Mitchell',  'Annual Physical',    60),   # Tuesday
        (date(2025, 3, 12), 'P-10678', 'Dr. James Okonkwo',   'Dermatology Consult',35),   # Wednesday
        (date(2025, 3, 13), 'P-11456', 'Dr. Priya Nair',      'General Checkup',    30),   # Thursday
        (date(2025, 3, 14), 'P-10234', 'Dr. Elena Vasquez',   'Follow-up',          20),   # Friday
        (date(2025, 3, 15), 'P-11678', 'Dr. Sarah Mitchell',  'Urgent Care',        30),   # Saturday (weekend)
        (date(2025, 3, 16), 'P-10789', 'Dr. James Okonkwo',   'Vaccination',        15),   # Sunday (weekend)
        (date(2025, 3, 17), 'P-11234', 'Dr. Priya Nair',      'Cardiology Consult', 60),   # Monday
        (date(2025, 3, 18), 'P-10456', 'Dr. Elena Vasquez',   'Blood Work Review',  45),   # Tuesday
        (date(2025, 3, 19), 'P-11890', 'Dr. Sarah Mitchell',  'Physical Therapy',   50),   # Wednesday
        (date(2025, 3, 20), 'P-10345', 'Dr. James Okonkwo',   'Post-Op Review',     40),   # Thursday
        (date(2025, 3, 21), 'P-11123', 'Dr. Priya Nair',      'Annual Physical',    60),   # Friday
        (date(2025, 3, 22), 'P-10567', 'Dr. Elena Vasquez',   'General Checkup',    30),   # Saturday (weekend)
    ]

    for r, (appt_date, pid, doctor, appt_type, duration) in enumerate(appointment_data, 2):
        ws.cell(row=r, column=1, value=appt_date).number_format = 'yyyy-mm-dd'
        ws.cell(row=r, column=2, value=pid)
        ws.cell(row=r, column=3, value=doctor)
        ws.cell(row=r, column=4, value=appt_type)
        ws.cell(row=r, column=5, value=duration)

    # Column widths for readability
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 24
    ws.column_dimensions['E'].width = 16

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
