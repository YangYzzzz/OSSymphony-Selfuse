"""
Initial Setup: Copy sheet task - create workbook with Template and Data sheets
Task ID: calc_ps_050
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_050'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Template ---
    ws_template = wb.active
    ws_template.title = 'Template'

    # Formatted table layout - Project Tracking Template
    # Title row (merged)
    ws_template.merge_cells('A1:F1')
    ws_template['A1'] = 'Project Status Report'
    ws_template['A1'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    ws_template['A1'].fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    ws_template['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Subtitle row
    ws_template.merge_cells('A2:F2')
    ws_template['A2'] = 'Quarter 1 - 2025'
    ws_template['A2'].font = Font(name='Arial', size=11, italic=True, color='2F5496')
    ws_template['A2'].alignment = Alignment(horizontal='center')

    # Headers in row 4
    headers = ['Project Name', 'Lead', 'Start Date', 'End Date', 'Budget ($)', 'Status']
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws_template.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    data = [
        ['Cloud Migration', 'Sarah Chen', '2025-01-10', '2025-03-31', 125000, 'In Progress'],
        ['Mobile App Redesign', 'Marcus Johnson', '2025-01-15', '2025-04-15', 85000, 'In Progress'],
        ['Data Warehouse Upgrade', 'Priya Patel', '2025-02-01', '2025-05-30', 210000, 'Planning'],
        ['Security Audit', 'James Wilson', '2025-01-05', '2025-02-28', 45000, 'Completed'],
        ['CRM Integration', 'Elena Rodriguez', '2025-02-15', '2025-06-30', 150000, 'Planning'],
        ['Network Overhaul', 'David Kim', '2025-03-01', '2025-07-15', 95000, 'Not Started'],
        ['AI Chatbot Pilot', 'Aisha Mohammed', '2025-01-20', '2025-04-20', 72000, 'In Progress'],
        ['ERP Module Rollout', 'Thomas Anderson', '2025-02-10', '2025-08-31', 340000, 'Planning'],
        ['Website Refresh', 'Lisa Chang', '2025-03-10', '2025-05-15', 58000, 'Not Started'],
        ['Compliance Framework', 'Robert Davis', '2025-01-25', '2025-03-25', 35000, 'In Progress'],
    ]

    data_font = Font(name='Arial', size=11)
    alt_fill = PatternFill(start_color='FFD9E2F3', end_color='FFD9E2F3', fill_type='solid')

    for r, row_data in enumerate(data, 5):
        for c, val in enumerate(row_data, 1):
            cell = ws_template.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = thin_border
            if r % 2 == 0:
                cell.fill = alt_fill
            if c == 5:  # Budget column
                cell.number_format = '$#,##0'
            if c in (3, 4):  # Date columns
                cell.alignment = Alignment(horizontal='center')

    # Summary row
    ws_template.cell(row=16, column=4, value='Total Budget:').font = Font(name='Arial', size=11, bold=True)
    total_cell = ws_template.cell(row=16, column=5, value='=SUM(E5:E14)')
    total_cell.font = Font(name='Arial', size=11, bold=True)
    total_cell.number_format = '$#,##0'
    total_cell.border = Border(top=Side(style='double', color='000000'), bottom=Side(style='double', color='000000'))

    # Column widths
    ws_template.column_dimensions['A'].width = 25
    ws_template.column_dimensions['B'].width = 20
    ws_template.column_dimensions['C'].width = 14
    ws_template.column_dimensions['D'].width = 14
    ws_template.column_dimensions['E'].width = 15
    ws_template.column_dimensions['F'].width = 14

    # Freeze panes
    ws_template.freeze_panes = 'A5'

    # --- Sheet 2: Data ---
    ws_data = wb.create_sheet('Data')

    # Reference data - Status options and department info
    ws_data['A1'] = 'Status Options'
    ws_data['A1'].font = Font(bold=True)
    statuses = ['Not Started', 'Planning', 'In Progress', 'Completed', 'On Hold', 'Cancelled']
    for i, s in enumerate(statuses, 2):
        ws_data.cell(row=i, column=1, value=s)

    ws_data['C1'] = 'Department'
    ws_data['D1'] = 'Head Count'
    ws_data['C1'].font = Font(bold=True)
    ws_data['D1'].font = Font(bold=True)
    departments = [
        ('Engineering', 45),
        ('Marketing', 22),
        ('Finance', 18),
        ('Operations', 31),
        ('HR', 12),
    ]
    for i, (dept, count) in enumerate(departments, 2):
        ws_data.cell(row=i, column=3, value=dept)
        ws_data.cell(row=i, column=4, value=count)

    ws_data.column_dimensions['A'].width = 16
    ws_data.column_dimensions['C'].width = 16
    ws_data.column_dimensions['D'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
