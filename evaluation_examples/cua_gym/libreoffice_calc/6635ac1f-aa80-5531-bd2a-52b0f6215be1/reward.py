"""
Reward Script: Copy 'Template' sheet within the same workbook, placed after 'Template', named 'Template (2)'
Task ID: calc_ps_050
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): 'Template (2)' sheet exists
  Component 2 (0.3): 'Template (2)' is positioned immediately after 'Template'
  Component 3 (0.4): 'Template (2)' cell data matches 'Template' (exact copy)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_050'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    sheet_names = wb.sheetnames
    print(f"INFO: Found sheets: {sheet_names}")

    # Component 1: 'Template (2)' sheet exists (0.3 points)
    # This FAILS on initial (only 'Template' and 'Data') and PASSES on golden
    try:
        if 'Template (2)' in sheet_names:
            print(f"PASS: Component 1 — 'Template (2)' sheet exists (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — 'Template (2)' sheet not found. Sheets: {sheet_names}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: 'Template (2)' is positioned immediately after 'Template' (0.3 points)
    # This FAILS on initial (sheet doesn't exist) and PASSES on golden
    try:
        if 'Template' in sheet_names and 'Template (2)' in sheet_names:
            idx_template = sheet_names.index('Template')
            idx_copy = sheet_names.index('Template (2)')
            if idx_copy == idx_template + 1:
                print(f"PASS: Component 2 — 'Template (2)' at index {idx_copy}, right after 'Template' at index {idx_template} (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — 'Template (2)' at index {idx_copy}, but 'Template' at index {idx_template}. Expected consecutive.")
        else:
            print(f"FAIL: Component 2 — Required sheets not found. Sheets: {sheet_names}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: 'Template (2)' is an exact data copy of 'Template' (0.4 points)
    # This FAILS on initial (sheet doesn't exist) and PASSES on golden
    try:
        if 'Template (2)' in sheet_names and 'Template' in sheet_names:
            ws_orig = wb['Template']
            ws_copy = wb['Template (2)']

            # Compare dimensions
            max_row = max(ws_orig.max_row, ws_copy.max_row)
            max_col = max(ws_orig.max_column, ws_copy.max_column)

            mismatches = 0
            total_cells = 0
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    v_orig = ws_orig.cell(row=row, column=col).value
                    v_copy = ws_copy.cell(row=row, column=col).value
                    if v_orig is not None or v_copy is not None:
                        total_cells += 1
                    if v_orig != v_copy:
                        mismatches += 1
                        if mismatches <= 3:
                            print(f"  MISMATCH at ({row},{col}): Template={v_orig}, Template (2)={v_copy}")

            if mismatches == 0 and total_cells > 0:
                print(f"PASS: Component 3 — All {total_cells} non-empty cells match between 'Template' and 'Template (2)' (0.4 pts)")
                total_score += 0.4
            elif total_cells == 0:
                print(f"FAIL: Component 3 — Both sheets appear empty")
            else:
                print(f"FAIL: Component 3 — {mismatches} cell mismatches out of {total_cells} non-empty cells")
        else:
            print(f"FAIL: Component 3 — Cannot compare: required sheets missing")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
