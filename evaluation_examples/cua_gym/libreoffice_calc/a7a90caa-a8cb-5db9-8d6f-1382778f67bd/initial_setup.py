"""
Initial Setup: Add date validation to the Deadline column
Task ID: calc_dop_validate_date_023
Domain: libreoffice_calc

Creates a Projects spreadsheet with 29 project records.
Column E (Deadline) contains a mix of dates: some in 2024, some in 2025, some in 2026.
NO data validation is present on column E in the initial file.
"""

import os
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_dop_validate_date_023'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Projects'

    # --- Headers ---
    headers = ['Project ID', 'Project Name', 'Manager', 'Start Date', 'Deadline', 'Budget']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # --- Project data with realistic content ---
    # Deadlines: mix of 2024 (past), 2025, and 2026 (future)
    projects = [
        ('PRJ-001', 'ERP System Migration',      'Sarah Chen',      date(2024, 1, 15),  date(2024, 6, 30),  125000),
        ('PRJ-002', 'Mobile App Redesign',        'Marcus Johnson',  date(2024, 2, 1),   date(2024, 11, 15), 87500),
        ('PRJ-003', 'Customer Portal Launch',     'Emily Rodriguez', date(2024, 3, 10),  date(2025, 3, 31),  210000),
        ('PRJ-004', 'Data Warehouse Upgrade',     'James Kim',       date(2024, 4, 5),   date(2025, 6, 15),  340000),
        ('PRJ-005', 'HR Platform Integration',    'Linda Patel',     date(2024, 5, 20),  date(2025, 9, 30),  95000),
        ('PRJ-006', 'Cybersecurity Audit',        'Tom Walker',      date(2024, 6, 1),   date(2024, 12, 20), 62000),
        ('PRJ-007', 'Cloud Infrastructure Setup', 'Anna Nguyen',     date(2024, 7, 15),  date(2025, 1, 31),  480000),
        ('PRJ-008', 'Inventory Management System','Robert Diaz',     date(2024, 8, 1),   date(2025, 4, 30),  153000),
        ('PRJ-009', 'Analytics Dashboard',        'Jessica Park',    date(2024, 9, 10),  date(2025, 7, 31),  72000),
        ('PRJ-010', 'API Gateway Modernization',  'Michael Scott',   date(2024, 10, 1),  date(2026, 2, 28),  198000),
        ('PRJ-011', 'Supply Chain Optimization',  'Olivia Turner',   date(2024, 11, 5),  date(2026, 5, 15),  275000),
        ('PRJ-012', 'Payment Gateway Upgrade',    'David Lee',       date(2024, 12, 1),  date(2025, 2, 28),  89000),
        ('PRJ-013', 'Network Infrastructure',     'Sarah Chen',      date(2025, 1, 15),  date(2025, 10, 31), 420000),
        ('PRJ-014', 'AI Chatbot Development',     'Marcus Johnson',  date(2025, 1, 20),  date(2025, 11, 30), 310000),
        ('PRJ-015', 'Business Intelligence Tools','Emily Rodriguez', date(2025, 2, 1),   date(2026, 3, 31),  165000),
        ('PRJ-016', 'Legacy System Retirement',   'James Kim',       date(2025, 2, 10),  date(2025, 8, 15),  78000),
        ('PRJ-017', 'DevOps Pipeline Setup',      'Linda Patel',     date(2025, 3, 1),   date(2025, 12, 31), 92000),
        ('PRJ-018', 'Compliance Management',      'Tom Walker',      date(2025, 3, 15),  date(2026, 6, 30),  145000),
        ('PRJ-019', 'E-Commerce Platform',        'Anna Nguyen',     date(2025, 4, 1),   date(2026, 1, 31),  520000),
        ('PRJ-020', 'Remote Work Tools',          'Robert Diaz',     date(2025, 4, 15),  date(2025, 9, 15),  58000),
        ('PRJ-021', 'Document Management System', 'Jessica Park',    date(2025, 5, 1),   date(2025, 12, 15), 113000),
        ('PRJ-022', 'Backup & Recovery Update',   'Michael Scott',   date(2025, 5, 10),  date(2025, 10, 1),  67000),
        ('PRJ-023', 'Digital Transformation',     'Olivia Turner',   date(2025, 6, 1),   date(2026, 9, 30),  780000),
        ('PRJ-024', 'CRM Implementation',         'David Lee',       date(2025, 6, 15),  date(2025, 11, 15), 235000),
        ('PRJ-025', 'Microservices Refactor',     'Sarah Chen',      date(2025, 7, 1),   date(2026, 4, 30),  195000),
        ('PRJ-026', 'VPN Infrastructure',         'Marcus Johnson',  date(2025, 7, 15),  date(2025, 12, 31), 44000),
        ('PRJ-027', 'Customer Feedback Portal',   'Emily Rodriguez', date(2025, 8, 1),   date(2024, 9, 30),  83000),
        ('PRJ-028', 'Automated Testing Framework','James Kim',       date(2025, 8, 20),  date(2025, 5, 31),  61000),
        ('PRJ-029', 'Sustainability Reporting',   'Linda Patel',     date(2025, 9, 1),   date(2024, 8, 15),  37000),
    ]

    for row_idx, (proj_id, name, manager, start_dt, deadline_dt, budget) in enumerate(projects, 2):
        ws.cell(row=row_idx, column=1, value=proj_id)
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=manager)
        # Store dates as actual date objects so LibreOffice formats them properly
        d_cell = ws.cell(row=row_idx, column=4, value=start_dt)
        d_cell.number_format = 'YYYY-MM-DD'
        e_cell = ws.cell(row=row_idx, column=5, value=deadline_dt)
        e_cell.number_format = 'YYYY-MM-DD'
        ws.cell(row=row_idx, column=6, value=budget)

    # --- Column widths ---
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
