"""
Initial Setup: Financial report spreadsheet with data spanning beyond print area.
Task ID: calc_nrv_014
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_014'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial Report"

    # --- Headers (Row 1) ---
    headers = ["Category", "Q1 Revenue", "Q2 Revenue", "Q3 Revenue", "Q4 Revenue", "Annual Total", "YoY Change %"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side = Side(style="thin", color="000000")
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # --- Data Rows (2-28): Department line items ---
    data = [
        ["Software Licenses", 45230, 48750, 52100, 55400, None, None],
        ["Cloud Infrastructure", 128500, 135200, 142800, 151000, None, None],
        ["Consulting Services", 67800, 71200, 68900, 74500, None, None],
        ["Hardware Sales", 34200, 31800, 38900, 42100, None, None],
        ["Support Contracts", 89100, 92400, 95700, 98200, None, None],
        ["Training Programs", 12400, 15800, 14200, 16900, None, None],
        ["Data Analytics", 56700, 61300, 64800, 69200, None, None],
        ["Security Solutions", 78900, 82100, 87400, 91500, None, None],
        ["Mobile Development", 23400, 27600, 31200, 35800, None, None],
        ["API Integration", 41200, 44500, 47800, 51200, None, None],
        ["DevOps Services", 33600, 36900, 39200, 42800, None, None],
        ["Quality Assurance", 28900, 31200, 33800, 36100, None, None],
        ["Project Management", 19800, 21400, 23100, 25600, None, None],
        ["UX Design", 15600, 17200, 19800, 22400, None, None],
        ["Technical Writing", 8900, 9400, 10200, 11800, None, None],
        ["Network Services", 62300, 65800, 69100, 72400, None, None],
        ["Database Admin", 44100, 47200, 50800, 53600, None, None],
        ["Systems Architecture", 37500, 39800, 42100, 45200, None, None],
        ["Compliance Auditing", 21600, 23400, 25800, 28100, None, None],
        ["Vendor Management", 16200, 17800, 19400, 21200, None, None],
        ["R&D Innovation", 95400, 102300, 108700, 115200, None, None],
        ["Customer Success", 27800, 30100, 33400, 36200, None, None],
        ["Marketing Tech", 18900, 21200, 24600, 27800, None, None],
        ["Legal Technology", 11200, 12800, 14500, 16200, None, None],
        ["Facilities IT", 7800, 8200, 8900, 9600, None, None],
        ["Executive Systems", 5400, 5800, 6200, 6700, None, None],
        ["Miscellaneous IT", 3200, 3600, 4100, 4500, None, None],
    ]

    currency_fmt = '$#,##0'
    pct_fmt = '0.0%'
    data_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = data_border
            if c == 1:
                cell.font = Font(name="Calibri", size=11)
            elif c <= 5:
                cell.number_format = currency_fmt
            elif c == 6:
                # Annual Total = sum of Q1-Q4
                cell.value = f"=SUM(B{r}:E{r})"
                cell.number_format = currency_fmt
            elif c == 7:
                # YoY Change % - simulate with a formula referencing previous year baseline
                # Use a simple calculation: (Q4-Q1)/Q1
                cell.value = f"=(E{r}-B{r})/B{r}"
                cell.number_format = pct_fmt

    # --- Row 29: Blank separator ---
    for c in range(1, 8):
        cell = ws.cell(row=29, column=c)
        cell.border = data_border

    # --- Row 30: Totals ---
    total_font = Font(name="Calibri", size=11, bold=True)
    total_fill = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")
    ws.cell(row=30, column=1, value="GRAND TOTAL").font = total_font
    ws.cell(row=30, column=1).fill = total_fill
    ws.cell(row=30, column=1).border = data_border

    for c in range(2, 6):
        cell = ws.cell(row=30, column=c)
        col_letter = chr(64 + c)
        cell.value = f"=SUM({col_letter}2:{col_letter}28)"
        cell.number_format = currency_fmt
        cell.font = total_font
        cell.fill = total_fill
        cell.border = data_border

    # Annual Total for grand total row
    cell = ws.cell(row=30, column=6)
    cell.value = "=SUM(F2:F28)"
    cell.number_format = currency_fmt
    cell.font = total_font
    cell.fill = total_fill
    cell.border = data_border

    # YoY Change for grand total row
    cell = ws.cell(row=30, column=7)
    cell.value = "=(E30-B30)/B30"
    cell.number_format = pct_fmt
    cell.font = total_font
    cell.fill = total_fill
    cell.border = data_border

    # --- Extra data beyond row 30 (rows 31-40): Notes and auxiliary data ---
    ws.cell(row=32, column=1, value="NOTES & AUXILIARY DATA (not for printing)")
    ws.cell(row=32, column=1).font = Font(name="Calibri", size=11, bold=True, italic=True)

    notes_data = [
        ["Report Period", "FY 2025", "", "", "", "", ""],
        ["Prepared By", "Sarah Chen, CFO", "", "", "", "", ""],
        ["Review Date", "2025-12-15", "", "", "", "", ""],
        ["Baseline Year", "FY 2024", "", "", "", "", ""],
        ["Currency", "USD", "", "", "", "", ""],
        ["Rounding", "Nearest dollar", "", "", "", "", ""],
        ["Audit Status", "Pending external review", "", "", "", "", ""],
        ["Distribution", "Board members only", "", "", "", "", ""],
    ]
    for r, row_data in enumerate(notes_data, 33):
        for c, val in enumerate(row_data, 1):
            if val:
                ws.cell(row=r, column=c, value=val)

    # --- Column widths ---
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 16

    # --- Row height for header ---
    ws.row_dimensions[1].height = 30

    # --- NO named ranges, NO print ranges (task requires agent to set these) ---

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
