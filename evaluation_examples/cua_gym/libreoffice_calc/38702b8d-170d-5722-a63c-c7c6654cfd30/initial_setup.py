"""
Initial Setup: Create a Grades spreadsheet with student scores.
Task ID: calc_lf_036
Domain: libreoffice_calc

The agent must: define named range TestScores for B2:B6, type range name
in D2, and build =AVERAGE(INDIRECT(D2)) in E2.
Initial state must NOT include: named range, D2 value, or E2 formula.
"""

import os
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_036'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Grades ---
    ws = wb.active
    ws.title = 'Grades'

    # Headers with light styling
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
    thin = Side(style="thin", color="000000")
    header_border = Border(bottom=thin)

    for col, header in enumerate(['Student', 'Score'], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center")

    # Student data (A2:B6)
    students = [
        ('Amy', 88),
        ('Bob', 76),
        ('Cal', 92),
        ('Dee', 81),
        ('Eve', 95),
    ]
    for r, (name, score) in enumerate(students, 2):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=score)

    # Column D & E: helper headers only -- agent fills D2 and E2
    ws['D1'] = 'Range Name'
    ws['D1'].font = Font(bold=True, size=11)
    ws['E1'] = 'Average'
    ws['E1'].font = Font(bold=True, size=11)

    # D2 and E2 are intentionally empty -- the agent must fill them
    # No named range defined -- the agent must create 'TestScores'

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 4
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
