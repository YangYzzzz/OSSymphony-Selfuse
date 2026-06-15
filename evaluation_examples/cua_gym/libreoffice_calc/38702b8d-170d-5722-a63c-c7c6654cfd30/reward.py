"""
Reward Script: Create named range 'TestScores' for B2:B6 and build a dynamic AVERAGE formula using INDIRECT.
Task ID: calc_lf_036
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Named range 'TestScores' defined for Grades!$B$2:$B$6
  Component 2 (0.2): D2 contains the text 'TestScores'
  Component 3 (0.4): E2 contains =AVERAGE(INDIRECT(D2)) formula
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_036'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists
    if 'Grades' not in wb.sheetnames:
        print("FAIL: Sheet 'Grades' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Grades']

    # Component 1: Named range 'TestScores' defined for Grades!$B$2:$B$6 (0.4 points)
    try:
        matching_names = [n for n in wb.defined_names if n.lower() == 'testscores']
        if matching_names:
            defn = wb.defined_names[matching_names[0]]
            attr = defn.attr_text
            # Normalize: strip quotes, check for Grades!$B$2:$B$6
            normalized = attr.replace("'", "").upper()
            if "GRADES!$B$2:$B$6" in normalized or "GRADES!B2:B6" in normalized:
                print(f"PASS: Component 1 — Named range 'TestScores' = {attr} (0.4 pts)")
                total_score += 0.4
            else:
                print(f"FAIL: Component 1 — Named range 'TestScores' found but value is '{attr}', expected Grades!$B$2:$B$6")
        else:
            print("FAIL: Component 1 — Named range 'TestScores' not defined")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: D2 contains 'TestScores' (0.2 points)
    try:
        d2_val = ws['D2'].value
        if d2_val is not None and str(d2_val).strip().lower() == 'testscores':
            print(f"PASS: Component 2 — D2 contains '{d2_val}' (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 2 — Expected D2='TestScores', found: {repr(d2_val)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: E2 contains =AVERAGE(INDIRECT(D2)) formula (0.4 points)
    try:
        e2_val = ws['E2'].value
        if e2_val is not None and isinstance(e2_val, str):
            # Normalize: remove spaces, uppercase for comparison
            normalized_formula = e2_val.upper().replace(" ", "")
            expected = "=AVERAGE(INDIRECT(D2))"
            expected_norm = expected.upper().replace(" ", "")
            if normalized_formula == expected_norm:
                print(f"PASS: Component 3 — E2 contains '{e2_val}' (0.4 pts)")
                total_score += 0.4
            else:
                print(f"FAIL: Component 3 — E2 formula is '{e2_val}', expected '{expected}'")
        else:
            print(f"FAIL: Component 3 — E2 is not a formula, found: {repr(e2_val)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
