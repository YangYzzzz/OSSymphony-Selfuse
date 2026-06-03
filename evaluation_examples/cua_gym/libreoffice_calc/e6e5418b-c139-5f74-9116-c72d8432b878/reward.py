"""
Reward Script: Fill in row totals and column totals using SUM formulas
Task ID: osworld_calc_fill_totals_004
Domain: libreoffice_calc
Scoring:
  Component 1: Row SUM formulas in G2:G21 (student totals)       — 0.50 pts (0.025 per cell)
  Component 2: Column SUM formulas in B22:F22 (subject totals)   — 0.40 pts (0.08 per cell)
  Component 3: Grand total SUM formula in G22                    — 0.10 pts
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_fill_totals_004'


def is_sum_formula(value):
    """Return True if value is a string that looks like a SUM formula."""
    if not isinstance(value, str):
        return False
    normalized = value.strip().upper().replace(' ', '')
    return normalized.startswith('=SUM(')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task:
      - Spreadsheet has students in A2:A21, subjects in B1:F1
      - Student scores in B2:F21
      - Column G (G2:G21) should have SUM row formulas (total per student)
      - Row 22 (B22:F22) should have SUM column formulas (total per subject)
      - G22 should have a SUM formula for the grand total
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the sheet exists (precondition gate — no score)
    if 'Grades' not in wb.sheetnames:
        print("CRITICAL: 'Grades' sheet not found.")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Grades']

    # ----------------------------------------------------------------
    # Component 1: Row SUM formulas in column G (G2:G21)
    #   Each student row should have a SUM formula covering their subject scores B-F.
    #   0.025 points per cell × 20 cells = 0.50 points total
    # ----------------------------------------------------------------
    row_formula_count = 0
    row_formula_missing = []

    try:
        for row in range(2, 22):  # rows 2-21 (20 students)
            cell_val = ws.cell(row=row, column=7).value  # column G
            if is_sum_formula(cell_val):
                row_formula_count += 1
            else:
                row_formula_missing.append(f"G{row}: {repr(cell_val)}")

        # Award 0.025 per correctly formulaed row
        if row_formula_count == 20:
            component1_score = 0.50
            print(f"PASS: Component 1 — All 20 row SUM formulas present in G2:G21 ({component1_score} pts)")
        elif row_formula_count > 0:
            component1_score = round(row_formula_count * 0.025, 4)
            print(f"PARTIAL: Component 1 — {row_formula_count}/20 row SUM formulas in G2:G21 ({component1_score} pts)")
            for detail in row_formula_missing[:5]:
                print(f"  Missing: {detail}")
        else:
            component1_score = 0.0
            print(f"FAIL: Component 1 — No SUM formulas found in G2:G21 (0.0 pts)")

        if component1_score > 0:
            total_score += component1_score

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ----------------------------------------------------------------
    # Component 2: Column SUM formulas in row 22 (B22:F22)
    #   Each subject column should have a SUM formula summing student rows 2-21.
    #   0.08 points per cell × 5 cells = 0.40 points total
    # ----------------------------------------------------------------
    col_formula_count = 0
    col_formula_missing = []

    try:
        col_letters = ['B', 'C', 'D', 'E', 'F']
        for idx, col_num in enumerate(range(2, 7)):  # columns B-F (cols 2-6)
            col_letter = col_letters[idx]
            cell_val = ws.cell(row=22, column=col_num).value
            if is_sum_formula(cell_val):
                col_formula_count += 1
            else:
                col_formula_missing.append(f"{col_letter}22: {repr(cell_val)}")

        # Award 0.08 per correctly formulaed column
        if col_formula_count == 5:
            component2_score = 0.40
            print(f"PASS: Component 2 — All 5 column SUM formulas present in B22:F22 ({component2_score} pts)")
        elif col_formula_count > 0:
            component2_score = round(col_formula_count * 0.08, 4)
            print(f"PARTIAL: Component 2 — {col_formula_count}/5 column SUM formulas in B22:F22 ({component2_score} pts)")
            for detail in col_formula_missing:
                print(f"  Missing: {detail}")
        else:
            component2_score = 0.0
            print(f"FAIL: Component 2 — No SUM formulas found in B22:F22 (0.0 pts)")

        if component2_score > 0:
            total_score += component2_score

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ----------------------------------------------------------------
    # Component 3: Grand total SUM formula in G22
    #   G22 should contain a SUM formula (typically =SUM(G2:G21)).
    #   0.10 points
    # ----------------------------------------------------------------
    try:
        g22_val = ws.cell(row=22, column=7).value
        if is_sum_formula(g22_val):
            print(f"PASS: Component 3 — Grand total SUM formula in G22: {g22_val} (0.1 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — Expected SUM formula in G22, found: {repr(g22_val)}")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
