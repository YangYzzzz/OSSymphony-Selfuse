"""
Initial Setup: Student Grade Matrix - Fill Totals Task
Task ID: osworld_calc_fill_totals_004
Domain: libreoffice_calc

Creates a student grade matrix with:
- Column A: Student names (rows 2-21)
- Columns B-F: Subject scores (Math, Science, English, History, Art)
- Column G header: "Total" — but cells G2:G21 are EMPTY (no formulas yet)
- Row 22 label: "Total" in A22 — but cells B22:F22 are EMPTY (no formulas yet)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_fill_totals_004'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Grades ---
    ws = wb.active
    ws.title = "Grades"

    # --- Header row (row 1) ---
    headers = ["Student", "Math", "Science", "English", "History", "Art", "Total"]
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    white_font = Font(bold=True, color="FFFFFFFF", size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Student data (rows 2-21) ---
    # Realistic student names and grade values (50-100 range, varied)
    students = [
        ("Sophia Martinez",    88,  92,  85,  79,  91),
        ("Liam Thompson",      74,  68,  80,  72,  65),
        ("Emma Nguyen",        95,  97,  93,  88,  90),
        ("Noah Patel",         61,  58,  70,  64,  55),
        ("Olivia Chen",        82,  86,  78,  83,  77),
        ("James Williams",     70,  73,  65,  68,  74),
        ("Ava Johnson",        90,  88,  94,  91,  87),
        ("Elijah Brown",       55,  62,  58,  60,  50),
        ("Isabella Davis",     78,  75,  82,  76,  80),
        ("Logan Garcia",       66,  70,  63,  71,  68),
        ("Mia Wilson",         93,  90,  96,  89,  92),
        ("Lucas Anderson",     72,  65,  69,  74,  67),
        ("Charlotte Lee",      84,  88,  81,  86,  83),
        ("Mason Taylor",       57,  53,  61,  56,  59),
        ("Amelia Harris",      87,  91,  84,  90,  88),
        ("Ethan Clark",        63,  67,  60,  65,  62),
        ("Harper Lewis",       79,  82,  76,  80,  75),
        ("Aiden Robinson",     96,  94,  98,  92,  95),
        ("Evelyn Walker",      68,  71,  65,  70,  66),
        ("Sebastian Hall",     81,  78,  83,  77,  82),
    ]

    # Style for data rows
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="FFBFBFBF")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Alternating row colors
    fill_light = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
    fill_white = PatternFill(fill_type=None)

    for r, (name, math, sci, eng, hist, art) in enumerate(students, 2):
        row_fill = fill_light if r % 2 == 0 else fill_white

        # Column A: Student name
        cell_a = ws.cell(row=r, column=1, value=name)
        cell_a.font = Font(size=11)
        cell_a.alignment = left_align
        cell_a.border = cell_border
        if r % 2 == 0:
            cell_a.fill = fill_light

        # Columns B-F: Subject scores
        for c, score in enumerate([math, sci, eng, hist, art], 2):
            cell = ws.cell(row=r, column=c, value=score)
            cell.font = Font(size=11)
            cell.alignment = center_align
            cell.border = cell_border
            if r % 2 == 0:
                cell.fill = fill_light

        # Column G: INTENTIONALLY EMPTY (no Total formula yet — task requires filling this)
        cell_g = ws.cell(row=r, column=7, value=None)
        cell_g.border = cell_border
        if r % 2 == 0:
            cell_g.fill = fill_light

    # --- Row 22: Totals row label (B22:F22 EMPTY — task requires filling this) ---
    total_fill = PatternFill(start_color="FFFFD966", end_color="FFFFD966", fill_type="solid")
    total_font = Font(bold=True, size=11)

    cell_a22 = ws.cell(row=22, column=1, value="Total")
    cell_a22.font = total_font
    cell_a22.fill = total_fill
    cell_a22.alignment = left_align
    cell_a22.border = cell_border

    for c in range(2, 8):  # B22:G22 — all empty, no formulas
        cell = ws.cell(row=22, column=c, value=None)
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = center_align
        cell.border = cell_border

    # --- Column widths ---
    ws.column_dimensions["A"].width = 22
    for col_letter in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_letter].width = 12

    # --- Row height for header ---
    ws.row_dimensions[1].height = 20

    # --- Freeze header row ---
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f"Initial file created: {OUTPUT}")

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print("GUI_READY: launched LibreOffice Calc with DISPLAY=:0")


create_initial()
