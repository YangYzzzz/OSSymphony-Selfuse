"""
Reward Script: Apply custom number format '#,##0;(#,##0)' to cells C2:C15
Task ID: calc_fmt_numfmt_negative_brackets_077
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): At least 7 of 14 cells in C2:C15 have the custom format '#,##0;(#,##0)'
  Component 2 (0.6): All 14 cells in C2:C15 have the custom format '#,##0;(#,##0)' AND
                     underlying values in C2:C15 are unchanged from the expected data
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmt_numfmt_negative_brackets_077'

# Expected number format string that must be applied to C2:C15
EXPECTED_FORMAT = '#,##0;(#,##0)'

# Expected data values in column C (rows 2–15), matching context ground truth
EXPECTED_VALUES = {
    2:  15000,
    3:  -8500,
    4:  22000,
    5:  -1200,
    6:  31400,
    7:  -14700,
    8:  18600,
    9:  -5300,
    10: 42800,
    11: -19200,
    12: 26500,
    13: -3700,
    14: 33100,
    15: -11400,
}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Cash Flow' sheet must exist
    if 'Cash Flow' not in wb.sheetnames:
        print(f"CRITICAL: Sheet 'Cash Flow' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Cash Flow']

    # Component 1: At least 7 of 14 cells in C2:C15 have the custom format '#,##0;(#,##0)'
    # This checks that the task was at least partially completed (FAILS on initial file).
    try:
        cells_with_format = 0
        cells_checked = 0
        for row in range(2, 16):
            cells_checked += 1
            fmt = ws.cell(row=row, column=3).number_format
            if fmt == EXPECTED_FORMAT:
                cells_with_format += 1

        if cells_with_format >= 7:
            print(f"PASS: Component 1 — {cells_with_format}/14 cells in C2:C15 have format '{EXPECTED_FORMAT}' (>= 7 required) (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — only {cells_with_format}/14 cells in C2:C15 have format '{EXPECTED_FORMAT}', need at least 7")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: ALL 14 cells in C2:C15 have the custom format '#,##0;(#,##0)'
    # AND underlying values remain unchanged from the expected cash flow data.
    # This ensures both completeness of formatting AND data integrity together.
    # FAILS on initial file (which has 'General' format), PASSES on golden file.
    try:
        format_failures = []
        value_failures = []

        for row in range(2, 16):
            fmt = ws.cell(row=row, column=3).number_format
            val = ws.cell(row=row, column=3).value
            expected_val = EXPECTED_VALUES[row]

            if fmt != EXPECTED_FORMAT:
                format_failures.append(f"C{row}: fmt={repr(fmt)}")

            if val != expected_val:
                value_failures.append(f"C{row}: val={val}, expected={expected_val}")

        if len(format_failures) == 0 and len(value_failures) == 0:
            print(f"PASS: Component 2 — all 14 cells in C2:C15 have format '{EXPECTED_FORMAT}' with unchanged values (0.6 pts)")
            total_score += 0.6
        else:
            if len(format_failures) > 0:
                print(f"FAIL: Component 2 — format errors: {format_failures}")
            if len(value_failures) > 0:
                print(f"FAIL: Component 2 — value errors (underlying data changed): {value_failures}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
