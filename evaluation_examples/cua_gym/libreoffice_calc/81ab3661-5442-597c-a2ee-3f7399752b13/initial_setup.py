"""
Initial Setup: Cash Flow spreadsheet with mixed positive/negative Net Change values
Task ID: calc_fmt_numfmt_negative_brackets_077
Domain: libreoffice_calc

Creates a 'Cash Flow' sheet with Period, Inflow, and Net Change columns.
Rows 2-15 contain realistic cash flow data.
Column C uses 'General' number format (task requires applying '#,##0;(#,##0)' to it).
"""

import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_numfmt_negative_brackets_077'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cash Flow'

    # --- Headers (Row 1) ---
    ws['A1'] = 'Period'
    ws['B1'] = 'Inflow'
    ws['C1'] = 'Net Change'

    # Bold headers
    for col in ['A', 'B', 'C']:
        ws[f'{col}1'].font = Font(bold=True)

    ws['A1'].alignment = Alignment(horizontal='center')
    ws['B1'].alignment = Alignment(horizontal='center')
    ws['C1'].alignment = Alignment(horizontal='center')

    # --- Data Rows 2-15 (realistic cash flow data, mixed +/-) ---
    # Period labels: Q1 2024 through Q4 2026 and some months
    data = [
        # Period,            Inflow,   Net Change
        ('Jan 2025',         85200,    15000),
        ('Feb 2025',         62400,    -8500),
        ('Mar 2025',         94700,    22000),
        ('Apr 2025',         78300,    -1200),
        ('May 2025',         103500,   31400),
        ('Jun 2025',         55800,    -14700),
        ('Jul 2025',         91200,    18600),
        ('Aug 2025',         67400,    -5300),
        ('Sep 2025',         112000,   42800),
        ('Oct 2025',         48900,    -19200),
        ('Nov 2025',         88600,    26500),
        ('Dec 2025',         76200,    -3700),
        ('Jan 2026',         95400,    33100),
        ('Feb 2026',         59100,    -11400),
    ]

    for r, (period, inflow, net_change) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=period)
        ws.cell(row=r, column=2, value=inflow)
        ws.cell(row=r, column=3, value=net_change)
        # Column C deliberately uses 'General' format — task is to apply custom format
        ws.cell(row=r, column=3).number_format = 'General'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Cash Flow')
    print(f'  Rows 2-15: 14 rows of cash flow data')
    print(f'  Column C: General number format (task target: #,##0;(#,##0))')


create_initial()
