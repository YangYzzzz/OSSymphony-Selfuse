"""
Initial Setup: Data validation on D2 for Q1 date checking
Task ID: calc_nrv_082
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_082'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # --- Headers ---
    headers = ['Expense', 'Amount', 'Category', 'Date (Q1 Only)']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Realistic expense data (10+ rows) ---
    # D2 is intentionally left empty (that's where validation goes)
    data = [
        ['Office Supplies', 142.50, 'Operations', None],          # Row 2 - D2 empty
        ['Client Lunch - Meridian Group', 87.30, 'Meals', '2026-01-15'],
        ['Software License - Jira', 450.00, 'Technology', '2026-02-01'],
        ['Travel - Chicago Conference', 1245.00, 'Travel', '2026-01-22'],
        ['Printer Ink Cartridges', 68.99, 'Operations', '2026-03-05'],
        ['Team Building Event', 325.00, 'HR', '2026-02-14'],
        ['Cloud Hosting - AWS', 892.47, 'Technology', '2026-01-31'],
        ['Marketing Brochures', 210.00, 'Marketing', '2026-03-10'],
        ['Courier Service - FedEx', 45.80, 'Logistics', '2026-02-20'],
        ['Professional Development - Workshop', 599.00, 'Training', '2026-01-08'],
        ['Ergonomic Keyboards (x5)', 375.00, 'Operations', '2026-03-18'],
        ['Quarterly Tax Filing Fee', 150.00, 'Finance', '2026-03-28'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 2 and val is not None:
                cell.number_format = '$#,##0.00'
            if c == 4 and val is not None:
                cell.number_format = 'yyyy-mm-dd'

    # --- Column widths ---
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 18

    # Freeze header row
    ws.freeze_panes = 'A2'

    # NO data validation on D2 - that's the task for the agent

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
