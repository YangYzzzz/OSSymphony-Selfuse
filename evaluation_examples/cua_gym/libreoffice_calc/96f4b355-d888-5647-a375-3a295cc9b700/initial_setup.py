"""
Initial Setup: Set up data validation for a budget sheet
Task ID: calc_nrv_067
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_067'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget"

    # --- Headers ---
    headers = {
        'A1': 'Dept',
        'B1': 'Project',
        'C1': 'Budget Amount',
        'D1': 'Category',
    }
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    for coord, val in headers.items():
        cell = ws[coord]
        cell.value = val
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Category reference list in G1:G8 ---
    categories = ['Capex', 'Opex', 'R&D', 'Marketing',
                  'Infrastructure', 'Personnel', 'Travel', 'Miscellaneous']
    for i, cat in enumerate(categories, 1):
        ws.cell(row=i, column=7, value=cat)

    # Label for the category list
    ws['F1'] = 'Categories:'
    ws['F1'].font = Font(italic=True, color="808080")

    # --- Dept and Project data (rows 2-20) for realism ---
    # C2:C20 and D2:D20 left EMPTY (no validation yet)
    dept_project_data = [
        ('Engineering', 'Cloud Migration Phase 2'),
        ('Marketing', 'Q3 Brand Campaign'),
        ('Finance', 'ERP System Upgrade'),
        ('Human Resources', 'Employee Wellness Program'),
        ('Operations', 'Warehouse Automation'),
        ('Engineering', 'Mobile App Redesign'),
        ('Sales', 'CRM Integration'),
        ('R&D', 'AI Prototype Development'),
        ('Marketing', 'Social Media Analytics Tool'),
        ('IT', 'Cybersecurity Enhancement'),
        ('Finance', 'Annual Audit Preparation'),
        ('Operations', 'Supply Chain Optimization'),
        ('Engineering', 'Data Pipeline Refactor'),
        ('Human Resources', 'Leadership Training'),
        ('Sales', 'Partner Portal Launch'),
        ('R&D', 'Patent Research Initiative'),
        ('IT', 'Network Infrastructure Refresh'),
        ('Marketing', 'Customer Retention Strategy'),
        ('Operations', 'Fleet Management System'),
    ]
    for r, (dept, project) in enumerate(dept_project_data, 2):
        ws.cell(row=r, column=1, value=dept)
        ws.cell(row=r, column=2, value=project)

    # --- Column widths ---
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 16

    # --- NO data validation on C or D columns ---

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
