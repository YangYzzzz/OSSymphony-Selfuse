"""
Reward Script: Edit comment on cell C7 of 'Audit' sheet to append resolution text
Task ID: calc_gg1_042
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Comment on C7 contains the appended resolution text
  Component 2 (0.3): Original comment text is preserved intact at the start
  Component 3 (0.3): Full combined comment matches expected text exactly
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_042'

ORIGINAL_TEXT = 'Discrepancy found: amount does not match invoice #4421'
APPENDED_TEXT = ' \u2013 Resolved by J. Smith on 2024-03-15'  # U+2013 en dash
EXPECTED_FULL = ORIGINAL_TEXT + APPENDED_TEXT


def persist_app_state(domain: str):
    """Attempt to save any unsaved changes in LibreOffice."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Audit' sheet must exist
    if 'Audit' not in wb.sheetnames:
        print(f"CRITICAL: 'Audit' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Audit']

    # Precondition: C7 must have a comment
    comment = ws['C7'].comment
    if comment is None:
        print("CRITICAL: Cell C7 has no comment at all")
        print("REWARD: 0.0")
        return 0.0

    comment_text = comment.text if comment.text else ''
    print(f"INFO: C7 comment text: {repr(comment_text)}")

    # Component 1: Comment contains the appended resolution text (0.4 points)
    # This checks that the resolution suffix was added. This FAILS on initial
    # (which only has the original text) and PASSES on golden.
    try:
        if APPENDED_TEXT in comment_text:
            print(f"PASS: Component 1 — Resolution text found in comment (0.4 pts)")
            total_score += 0.4
        else:
            # Also check with a regular hyphen in case of dash variation
            alt_appended = ' - Resolved by J. Smith on 2024-03-15'
            if alt_appended in comment_text and 'Resolved by J. Smith on 2024-03-15' in comment_text:
                print(f"PASS: Component 1 — Resolution text found (alt dash) (0.4 pts)")
                total_score += 0.4
            else:
                print(f"FAIL: Component 1 — Resolution text not found. Expected substring: {repr(APPENDED_TEXT)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Original comment text is preserved at the start (0.3 points)
    # The original finding text must still be present and at the beginning.
    # This FAILS on initial because the full expected text (original + appended)
    # is not present — the initial only has the original text without the appended part.
    # We check that the comment STARTS with the original text AND contains more after it.
    try:
        starts_with_original = comment_text.startswith(ORIGINAL_TEXT)
        has_more_than_original = len(comment_text) > len(ORIGINAL_TEXT)
        if starts_with_original and has_more_than_original:
            print(f"PASS: Component 2 — Original text preserved and comment was extended (0.3 pts)")
            total_score += 0.3
        elif starts_with_original and not has_more_than_original:
            print(f"FAIL: Component 2 — Original text present but nothing was appended")
        else:
            print(f"FAIL: Component 2 — Comment does not start with original text")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Full combined comment matches expected text exactly (0.3 points)
    # This checks the precise final text. FAILS on initial, PASSES on golden.
    try:
        if comment_text == EXPECTED_FULL:
            print(f"PASS: Component 3 — Full comment text matches exactly (0.3 pts)")
            total_score += 0.3
        else:
            # Allow minor variation: regular dash instead of en dash
            alt_expected = ORIGINAL_TEXT + ' - Resolved by J. Smith on 2024-03-15'
            if comment_text == alt_expected:
                print(f"PASS: Component 3 — Full comment text matches (alt dash) (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 3 — Full text mismatch")
                print(f"  Expected: {repr(EXPECTED_FULL)}")
                print(f"  Got:      {repr(comment_text)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
