"""
Initial Setup: Edit comment on cell C7 of the 'Audit' sheet
Task ID: calc_gg1_042
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_042'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Audit ---
    ws = wb.active
    ws.title = 'Audit'

    # Headers
    headers = ['Finding ID', 'Date', 'Category', 'Description', 'Amount', 'Status', 'Assigned To']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows - realistic audit findings
    data = [
        ['AF-2024-001', '2024-01-10', 'Revenue', 'Unrecorded sales transaction for Q4 client engagement', 12500.00, 'Open', 'M. Rivera'],
        ['AF-2024-002', '2024-01-15', 'Expense', 'Missing receipts for travel reimbursement batch #221', 3420.75, 'Resolved', 'K. Patel'],
        ['AF-2024-003', '2024-01-22', 'Payroll', 'Overtime calculation error for night-shift staff', 8750.00, 'Under Review', 'L. Nguyen'],
        ['AF-2024-004', '2024-02-05', 'Inventory', 'Physical count variance in warehouse section B3', 15200.00, 'Open', 'T. Harrison'],
        ['AF-2024-005', '2024-02-12', 'Compliance', 'Late filing of quarterly tax remittance', 0.00, 'Resolved', 'A. Gomez'],
        ['AF-2024-006', '2024-02-20', 'Revenue', 'Discrepancy found: amount does not match invoice #4421', 4850.00, 'Open', 'J. Smith'],
        ['AF-2024-007', '2024-03-01', 'Expense', 'Duplicate vendor payment to Greenfield Supplies', 6300.00, 'Under Review', 'M. Rivera'],
        ['AF-2024-008', '2024-03-05', 'Payroll', 'Incorrect benefit deduction for new hires in March', 1125.50, 'Open', 'K. Patel'],
        ['AF-2024-009', '2024-03-10', 'Inventory', 'Obsolete stock not written off per policy ICP-09', 22000.00, 'Open', 'L. Nguyen'],
        ['AF-2024-010', '2024-03-12', 'Compliance', 'Missing approval signature on purchase order #7892', 9400.00, 'Under Review', 'T. Harrison'],
        ['AF-2024-011', '2024-03-18', 'Revenue', 'Revenue recognition timing issue for contract #C-445', 31000.00, 'Open', 'A. Gomez'],
        ['AF-2024-012', '2024-03-25', 'Expense', 'Unauthorized credit card charge on corporate account', 2150.00, 'Under Review', 'J. Smith'],
    ]

    date_format = 'yyyy-mm-dd'
    currency_format = '$#,##0.00'
    data_font = Font(name='Calibri', size=11)

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = thin_border
            if c == 2:  # Date column
                cell.number_format = date_format
            elif c == 5:  # Amount column
                cell.number_format = currency_format
                cell.alignment = Alignment(horizontal='right')
            elif c == 1:  # Finding ID
                cell.alignment = Alignment(horizontal='center')

    # Set column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 55
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 16

    # Add comment to C7 (row 7, column 3) - this is the cell in the Category column of row 7
    # Row 7 = data row 6 (AF-2024-006), Category = "Revenue"
    comment_text = 'Discrepancy found: amount does not match invoice #4421'
    comment = Comment(comment_text, 'Audit System')
    ws['C7'].comment = comment

    # Freeze the header row
    ws.freeze_panes = 'A2'

    # --- Sheet 2: Summary ---
    ws2 = wb.create_sheet('Summary')
    ws2['A1'] = 'Audit Findings Summary'
    ws2['A1'].font = Font(name='Calibri', size=14, bold=True)

    summary_headers = ['Category', 'Total Findings', 'Total Amount', 'Open', 'Resolved', 'Under Review']
    for col, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = Font(name='Calibri', size=11, bold=True)
        cell.fill = PatternFill(start_color='FFD9E2F3', end_color='FFD9E2F3', fill_type='solid')

    summary_data = [
        ['Revenue', 3, 48350.00, 2, 0, 1],
        ['Expense', 3, 11870.75, 0, 1, 2],
        ['Payroll', 2, 9875.50, 1, 0, 1],
        ['Inventory', 2, 37200.00, 2, 0, 0],
        ['Compliance', 2, 9400.00, 0, 1, 1],
    ]

    for r, row_data in enumerate(summary_data, 4):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if c == 3:
                cell.number_format = currency_format

    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 16
    ws2.column_dimensions['D'].width = 10
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
