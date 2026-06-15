"""
Reward Script: Hide all rows where Supplier OR Delivery Date is 'N/A'
Task ID: osworld_calc_hide_rows_na_006
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5 pts): All rows with Supplier == 'N/A' are hidden
  Component 2 (0.5 pts): All rows with Delivery Date == 'N/A' are hidden
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_hide_rows_na_006'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task: Hide all rows where Supplier (col B) == 'N/A' OR Delivery Date (col D) == 'N/A'.
    Rows must remain in the spreadsheet (not deleted), just hidden.

    Scoring breakdown:
      - Component 1 (0.5): Every data row with Supplier == 'N/A' is hidden
      - Component 2 (0.5): Every data row with Delivery Date == 'N/A' is hidden

    Both components also verify that rows with VALID values in both columns
    are NOT hidden (they are sub-conditions ensuring correctness).
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: Check that the expected sheet exists
    if 'Supply Chain' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Supply Chain' not found. Cannot verify task.")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Supply Chain']

    # Precondition gate: Verify expected header row
    headers = [ws.cell(row=1, column=c).value for c in range(1, 6)]
    expected_headers = ['PO ID', 'Supplier', 'Product', 'Delivery Date', 'Amount']
    if headers != expected_headers:
        print(f"CRITICAL: Unexpected headers: {headers}. Expected: {expected_headers}")
        print("REWARD: 0.0")
        return 0.0

    # Collect all data rows (skip header row 1)
    # Columns: 1=PO ID, 2=Supplier, 3=Product, 4=Delivery Date, 5=Amount
    na_supplier_rows = []      # rows where Supplier == 'N/A'
    na_delivery_rows = []      # rows where Delivery Date == 'N/A'
    valid_rows = []             # rows where BOTH Supplier and Delivery Date are valid

    for row_num in range(2, ws.max_row + 1):
        supplier_val = ws.cell(row=row_num, column=2).value
        delivery_val = ws.cell(row=row_num, column=4).value

        supplier_is_na = (str(supplier_val).strip() == 'N/A') if supplier_val is not None else False
        delivery_is_na = (str(delivery_val).strip() == 'N/A') if delivery_val is not None else False

        if supplier_is_na:
            na_supplier_rows.append(row_num)
        if delivery_is_na:
            na_delivery_rows.append(row_num)
        if not supplier_is_na and not delivery_is_na:
            valid_rows.append(row_num)

    print(f"Found {len(na_supplier_rows)} rows with N/A Supplier: rows {na_supplier_rows}")
    print(f"Found {len(na_delivery_rows)} rows with N/A Delivery Date: rows {na_delivery_rows}")
    print(f"Found {len(valid_rows)} rows with both valid (should remain visible): rows {valid_rows}")
    print()

    # Component 1: All rows with Supplier == 'N/A' are hidden (0.5 points)
    # This fails on initial_env (all rows visible) and passes on golden_env
    try:
        supplier_na_not_hidden = [r for r in na_supplier_rows if not ws.row_dimensions[r].hidden]
        valid_incorrectly_hidden = [r for r in valid_rows if ws.row_dimensions[r].hidden]

        if len(na_supplier_rows) > 0 and len(supplier_na_not_hidden) == 0 and len(valid_incorrectly_hidden) == 0:
            print(f"PASS: Component 1 — All {len(na_supplier_rows)} rows with Supplier=N/A are correctly hidden (0.5 pts)")
            total_score += 0.5
        elif len(na_supplier_rows) == 0:
            print("FAIL: Component 1 — No N/A supplier rows found (unexpected data state)")
        elif len(supplier_na_not_hidden) > 0:
            print(f"FAIL: Component 1 — {len(supplier_na_not_hidden)} rows with Supplier=N/A are still visible: rows {supplier_na_not_hidden}")
        else:
            print(f"FAIL: Component 1 — {len(valid_incorrectly_hidden)} valid rows are incorrectly hidden (false positives): rows {valid_incorrectly_hidden}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All rows with Delivery Date == 'N/A' are hidden (0.5 points)
    # This fails on initial_env (all rows visible) and passes on golden_env
    try:
        delivery_na_not_hidden = [r for r in na_delivery_rows if not ws.row_dimensions[r].hidden]

        if len(na_delivery_rows) > 0 and len(delivery_na_not_hidden) == 0:
            print(f"PASS: Component 2 — All {len(na_delivery_rows)} rows with Delivery Date=N/A are correctly hidden (0.5 pts)")
            total_score += 0.5
        elif len(na_delivery_rows) == 0:
            print("FAIL: Component 2 — No N/A delivery date rows found (unexpected data state)")
        else:
            print(f"FAIL: Component 2 — {len(delivery_na_not_hidden)} rows with Delivery Date=N/A are still visible: rows {delivery_na_not_hidden}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the VM environment
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
