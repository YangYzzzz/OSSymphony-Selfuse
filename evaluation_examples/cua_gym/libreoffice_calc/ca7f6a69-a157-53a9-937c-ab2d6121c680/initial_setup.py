"""
Initial Setup: Supply chain spreadsheet with N/A values in Supplier and Delivery Date columns
Task ID: osworld_calc_hide_rows_na_006
Domain: libreoffice_calc

Creates a spreadsheet with supply chain records. Some rows have 'N/A' in Supplier,
some in Delivery Date, some in both, and some with fully valid values.
The agent task is to hide all rows where either column is 'N/A'.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_hide_rows_na_006'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Supply Chain"

    # --- Header Row ---
    headers = ['PO ID', 'Supplier', 'Product', 'Delivery Date', 'Amount']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 20
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 14

    # --- Data rows ---
    # Mix of: valid rows, N/A in Supplier, N/A in Delivery Date, N/A in both
    # Rows with N/A in Supplier OR Delivery Date should be hidden by the agent
    data = [
        # PO ID,    Supplier,              Product,                 Delivery Date,  Amount
        # Row 2 - valid
        ['PO-10021', 'Acme Industrial',    'Steel Beams (20ft)',    '2025-04-08',   18750.00],
        # Row 3 - N/A in Supplier
        ['PO-10022', 'N/A',                'Copper Wire (500m)',    '2025-04-12',    4320.50],
        # Row 4 - valid
        ['PO-10023', 'BrightTech Mfg',     'LED Panels (100 units)', '2025-04-15',  9870.00],
        # Row 5 - N/A in Delivery Date
        ['PO-10024', 'Sunrise Logistics',  'Polymer Resin (50kg)',   'N/A',          2640.00],
        # Row 6 - valid
        ['PO-10025', 'GlobalParts Co.',    'Hydraulic Pumps (8)',    '2025-04-20',  33600.00],
        # Row 7 - N/A in both
        ['PO-10026', 'N/A',                'Circuit Boards (200)',   'N/A',         15200.00],
        # Row 8 - valid
        ['PO-10027', 'Meridian Supplies',  'Industrial Solvents (10L)', '2025-04-22', 1875.60],
        # Row 9 - N/A in Supplier
        ['PO-10028', 'N/A',                'Titanium Rods (12)',     '2025-04-25',  27450.00],
        # Row 10 - valid
        ['PO-10029', 'Vertex Components',  'Servo Motors (15)',      '2025-04-28',  11220.00],
        # Row 11 - N/A in Delivery Date
        ['PO-10030', 'EastCoast Metals',   'Aluminum Sheets (30)',   'N/A',          8910.00],
        # Row 12 - valid
        ['PO-10031', 'Northern Plastics',  'PVC Piping (100m)',      '2025-05-03',   3250.00],
        # Row 13 - N/A in Delivery Date
        ['PO-10032', 'Pacific Fasteners',  'Stainless Bolts (1000)', 'N/A',          1650.00],
        # Row 14 - valid
        ['PO-10033', 'Summit Engineering', 'Gear Assemblies (6)',    '2025-05-06',  22440.00],
        # Row 15 - N/A in Supplier
        ['PO-10034', 'N/A',                'Rubber Gaskets (500)',   '2025-05-09',    980.00],
        # Row 16 - valid
        ['PO-10035', 'Cascade Materials',  'Fiberglass Rolls (20)',  '2025-05-12',   6720.00],
        # Row 17 - N/A in both
        ['PO-10036', 'N/A',                'Electronic Sensors (50)', 'N/A',         7500.00],
        # Row 18 - valid
        ['PO-10037', 'Pioneer Industries', 'Carbon Fiber Sheets (10)', '2025-05-15', 14300.00],
        # Row 19 - N/A in Delivery Date
        ['PO-10038', 'Bluestone Logistics', 'Industrial Filters (25)', 'N/A',         3375.00],
        # Row 20 - valid
        ['PO-10039', 'Ironclad Supplies',  'Steel Cables (500m)',    '2025-05-19',  10850.00],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # NOTE: No rows hidden in initial state. The agent must hide N/A rows.
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
