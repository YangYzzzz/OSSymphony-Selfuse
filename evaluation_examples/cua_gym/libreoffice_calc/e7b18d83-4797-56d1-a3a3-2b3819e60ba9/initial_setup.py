"""
Initial Setup: Create workbook with Sales, Expenses, Summary sheets for BatchRename macro task
Task ID: calc_mcp_028
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_028'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Shared styles ---
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    def style_header(ws, headers, row=1):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    # ============================================================
    # Sheet 1: Sales
    # ============================================================
    ws_sales = wb.active
    ws_sales.title = "Sales"

    sales_headers = ["Region", "Product", "Q1 Revenue", "Q2 Revenue", "Q3 Revenue", "Q4 Revenue", "Annual Total"]
    style_header(ws_sales, sales_headers)

    sales_data = [
        ["North America", "Widget Pro", 145200, 162800, 178400, 195600, None],
        ["North America", "GadgetX", 89500, 94200, 101300, 108700, None],
        ["Europe", "Widget Pro", 112300, 125600, 131900, 142500, None],
        ["Europe", "GadgetX", 67800, 72400, 78900, 83200, None],
        ["Asia Pacific", "Widget Pro", 98700, 115400, 128600, 139200, None],
        ["Asia Pacific", "GadgetX", 54300, 61800, 69200, 75600, None],
        ["Latin America", "Widget Pro", 43200, 48900, 55100, 62300, None],
        ["Latin America", "GadgetX", 28100, 32500, 37800, 41600, None],
        ["Middle East", "Widget Pro", 31500, 36200, 42800, 49100, None],
        ["Middle East", "GadgetX", 19800, 23400, 27600, 31200, None],
        ["Africa", "Widget Pro", 15600, 18900, 22300, 26700, None],
        ["Africa", "GadgetX", 9400, 11200, 13800, 16500, None],
    ]

    for r, row_data in enumerate(sales_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_sales.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c >= 3 and val is not None:
                cell.number_format = '$#,##0'

    # Annual Total formulas
    for r in range(2, 14):
        cell = ws_sales.cell(row=r, column=7, value=f'=SUM(C{r}:F{r})')
        cell.border = thin_border
        cell.number_format = '$#,##0'

    # Column widths
    ws_sales.column_dimensions["A"].width = 16
    ws_sales.column_dimensions["B"].width = 14
    for col_letter in ["C", "D", "E", "F", "G"]:
        ws_sales.column_dimensions[col_letter].width = 14

    ws_sales.freeze_panes = "A2"

    # ============================================================
    # Sheet 2: Expenses
    # ============================================================
    ws_expenses = wb.create_sheet("Expenses")

    expense_headers = ["Department", "Category", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "H1 Total"]
    style_header(ws_expenses, expense_headers)

    expense_data = [
        ["Engineering", "Salaries", 185000, 185000, 185000, 190000, 190000, 190000],
        ["Engineering", "Equipment", 12500, 8200, 15600, 9800, 22400, 7300],
        ["Engineering", "Software Licenses", 4500, 4500, 4500, 4500, 4500, 4500],
        ["Marketing", "Salaries", 95000, 95000, 95000, 98000, 98000, 98000],
        ["Marketing", "Advertising", 32000, 45000, 28000, 51000, 38000, 42000],
        ["Marketing", "Events", 8500, 0, 15200, 0, 22800, 0],
        ["Operations", "Salaries", 120000, 120000, 120000, 123000, 123000, 123000],
        ["Operations", "Facilities", 18500, 18500, 18500, 18500, 18500, 18500],
        ["Operations", "Utilities", 6200, 5800, 5400, 6800, 7200, 7800],
        ["HR", "Salaries", 68000, 68000, 68000, 70000, 70000, 70000],
        ["HR", "Recruitment", 15000, 22000, 8500, 18000, 12000, 25000],
        ["HR", "Training", 5200, 3800, 9600, 4200, 7500, 6100],
    ]

    for r, row_data in enumerate(expense_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_expenses.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c >= 3:
                cell.number_format = '$#,##0'

    # H1 Total formulas
    for r in range(2, 14):
        cell = ws_expenses.cell(row=r, column=9, value=f'=SUM(C{r}:H{r})')
        cell.border = thin_border
        cell.number_format = '$#,##0'

    ws_expenses.column_dimensions["A"].width = 14
    ws_expenses.column_dimensions["B"].width = 18
    for col_letter in ["C", "D", "E", "F", "G", "H", "I"]:
        ws_expenses.column_dimensions[col_letter].width = 12

    ws_expenses.freeze_panes = "A2"

    # ============================================================
    # Sheet 3: Summary
    # ============================================================
    ws_summary = wb.create_sheet("Summary")

    summary_headers = ["Metric", "Value", "Notes"]
    style_header(ws_summary, summary_headers)

    summary_data = [
        ["Total Annual Revenue", 2458700, "Combined all regions"],
        ["Total H1 Expenses", 3891500, "All departments"],
        ["Headcount", 247, "As of June 2024"],
        ["Revenue per Employee", 9954.25, "Annual basis"],
        ["Top Region", "North America", "By total revenue"],
        ["Top Product", "Widget Pro", "By total revenue"],
        ["YoY Growth Rate", 0.124, "Compared to FY2023"],
        ["Customer Satisfaction", 0.871, "Survey average"],
        ["Market Share", 0.183, "Industry estimate"],
        ["Operating Margin", 0.215, "Before tax"],
    ]

    for r, row_data in enumerate(summary_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_summary.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 2 and isinstance(val, float) and val < 1:
                cell.number_format = '0.0%'
            elif c == 2 and isinstance(val, (int, float)) and val > 100:
                cell.number_format = '$#,##0.00'

    ws_summary.column_dimensions["A"].width = 24
    ws_summary.column_dimensions["B"].width = 18
    ws_summary.column_dimensions["C"].width = 24

    # Save
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
