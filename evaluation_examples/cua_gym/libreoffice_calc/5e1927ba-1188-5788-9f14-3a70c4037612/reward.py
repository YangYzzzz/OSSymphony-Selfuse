"""
Reward Script: Build FTE Budget Planning Model
Task ID: calc_hr_fte_budget_planning_051
Domain: libreoffice_calc
Scoring:
  - Component 1: D2:D9 formulas =B*C for total salary cost (0.30 pts)
  - Component 2: F2:F9 and G2:G9 formulas for variance (0.30 pts)
  - Component 3: Number formatting in D, F, G columns (0.20 pts)
  - Component 4: Row 10 total row with SUM formulas and bold (0.10 pts)
  - Component 5: Conditional formatting on A2:G9 (formula >0.1, orange fill) (0.10 pts)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_hr_fte_budget_planning_051'


def verify_task(file_path):
    """
    Verify FTE budget planning model completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the 'FTE Plan' sheet exists
    if 'FTE Plan' not in wb.sheetnames:
        print("CRITICAL: Sheet 'FTE Plan' not found.")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['FTE Plan']

    # -------------------------------------------------------------------------
    # Component 1: D2:D9 contain formula =Bx*Cx (total salary cost) (0.30 pts)
    # These cells are EMPTY in the initial file, so any presence of =B*C formula
    # is a task-introduced change.
    # -------------------------------------------------------------------------
    try:
        d_formula_count = 0
        d_formula_correct = 0
        for row in range(2, 10):  # rows 2 to 9
            val = ws.cell(row=row, column=4).value  # Column D
            if val is not None and isinstance(val, str):
                d_formula_count += 1
                # Check pattern like =B2*C2 (row-relative)
                expected_formula = f'=B{row}*C{row}'
                if val.upper().replace(' ', '') == expected_formula.upper().replace(' ', ''):
                    d_formula_correct += 1
        if d_formula_correct == 8:
            print(f"PASS: Component 1 — All 8 D-column formulas (=Bx*Cx) present and correct (0.30 pts)")
            total_score += 0.30
        elif d_formula_correct >= 4:
            partial = round(0.30 * d_formula_correct / 8, 4)
            print(f"PARTIAL: Component 1 — {d_formula_correct}/8 D-column formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {d_formula_correct}/8 D-column formulas correct (expected =Bx*Cx)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: F2:F9 contain formula =Dx-Ex AND G2:G9 contain formula =Fx/Ex
    # Both columns empty in initial file — task-introduced changes.
    # (0.30 pts)
    # -------------------------------------------------------------------------
    try:
        f_correct = 0
        g_correct = 0
        for row in range(2, 10):  # rows 2 to 9
            f_val = ws.cell(row=row, column=6).value  # Column F
            g_val = ws.cell(row=row, column=7).value  # Column G
            expected_f = f'=D{row}-E{row}'
            expected_g = f'=F{row}/E{row}'
            if f_val is not None and isinstance(f_val, str):
                if f_val.upper().replace(' ', '') == expected_f.upper().replace(' ', ''):
                    f_correct += 1
            if g_val is not None and isinstance(g_val, str):
                if g_val.upper().replace(' ', '') == expected_g.upper().replace(' ', ''):
                    g_correct += 1

        if f_correct == 8 and g_correct == 8:
            print(f"PASS: Component 2 — All F2:F9 (=Dx-Ex) and G2:G9 (=Fx/Ex) formulas correct (0.30 pts)")
            total_score += 0.30
        elif f_correct + g_correct >= 8:
            partial = round(0.30 * (f_correct + g_correct) / 16, 4)
            print(f"PARTIAL: Component 2 — F correct: {f_correct}/8, G correct: {g_correct}/8 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — F correct: {f_correct}/8 (=Dx-Ex), G correct: {g_correct}/8 (=Fx/Ex)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Number formatting
    #   - D2:D9 and F2:F9 formatted as $#,##0
    #   - G2:G9 formatted as 0.0%
    # Initial file has no values in D, F, G, so any format there is task-introduced.
    # (0.20 pts)
    # -------------------------------------------------------------------------
    try:
        d_format_ok = 0
        f_format_ok = 0
        g_format_ok = 0
        for row in range(2, 10):
            d_fmt = ws.cell(row=row, column=4).number_format
            f_fmt = ws.cell(row=row, column=6).number_format
            g_fmt = ws.cell(row=row, column=7).number_format
            if '$#,##0' in d_fmt:
                d_format_ok += 1
            if '$#,##0' in f_fmt:
                f_format_ok += 1
            if '0.0%' in g_fmt or '0.00%' in g_fmt:
                g_format_ok += 1

        format_score = 0.0
        if d_format_ok == 8:
            format_score += 0.07
            print(f"PASS: Component 3a — D2:D9 formatted as $#,##0 (0.07 pts)")
        else:
            print(f"FAIL: Component 3a — D column: {d_format_ok}/8 cells formatted as $#,##0")

        if f_format_ok == 8:
            format_score += 0.07
            print(f"PASS: Component 3b — F2:F9 formatted as $#,##0 (0.07 pts)")
        else:
            print(f"FAIL: Component 3b — F column: {f_format_ok}/8 cells formatted as $#,##0")

        if g_format_ok == 8:
            format_score += 0.06
            print(f"PASS: Component 3c — G2:G9 formatted as 0.0% (0.06 pts)")
        else:
            print(f"FAIL: Component 3c — G column: {g_format_ok}/8 cells formatted as 0.0%")

        total_score += format_score
        if format_score == 0.20:
            print(f"PASS: Component 3 — All number formats correct (0.20 pts)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: Row 10 totals
    #   - A10 = 'Total' (bold)
    #   - D10 = =SUM(D2:D9) (bold, $#,##0 format)
    #   - E10 = =SUM(E2:E9) (bold, $#,##0 format)
    #   - F10 = =SUM(F2:F9) (bold, $#,##0 format)
    # Row 10 is entirely absent from initial file — task-introduced change.
    # (0.10 pts)
    # -------------------------------------------------------------------------
    try:
        a10 = ws.cell(row=10, column=1)
        d10 = ws.cell(row=10, column=4)
        e10 = ws.cell(row=10, column=5)
        f10 = ws.cell(row=10, column=6)

        a10_ok = (a10.value == 'Total' and a10.font.bold)
        d10_formula = d10.value
        e10_formula = e10.value
        f10_formula = f10.value

        d10_ok = (
            isinstance(d10_formula, str) and
            d10_formula.upper().replace(' ', '') == '=SUM(D2:D9)' and
            d10.font.bold and
            '$#,##0' in d10.number_format
        )
        e10_ok = (
            isinstance(e10_formula, str) and
            e10_formula.upper().replace(' ', '') == '=SUM(E2:E9)' and
            e10.font.bold and
            '$#,##0' in e10.number_format
        )
        f10_ok = (
            isinstance(f10_formula, str) and
            f10_formula.upper().replace(' ', '') == '=SUM(F2:F9)' and
            f10.font.bold and
            '$#,##0' in f10.number_format
        )

        if a10_ok and d10_ok and e10_ok and f10_ok:
            print(f"PASS: Component 4 — Row 10 totals (A10='Total' bold, D10/E10/F10 SUM formulas bold + $#,##0) (0.10 pts)")
            total_score += 0.10
        else:
            details = []
            if not a10_ok:
                details.append(f"A10: value={repr(a10.value)}, bold={a10.font.bold}")
            if not d10_ok:
                details.append(f"D10: value={repr(d10_formula)}, bold={d10.font.bold}, fmt={d10.number_format!r}")
            if not e10_ok:
                details.append(f"E10: value={repr(e10_formula)}, bold={e10.font.bold}, fmt={e10.number_format!r}")
            if not f10_ok:
                details.append(f"F10: value={repr(f10_formula)}, bold={f10.font.bold}, fmt={f10.number_format!r}")
            print(f"FAIL: Component 4 — Row 10 totals issues: {'; '.join(details)}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -------------------------------------------------------------------------
    # Component 5: Conditional formatting on A2:G9
    #   - formula: $G2>0.1 (highlight rows where G column > 10%)
    #   - fill: orange background (FFFF6600)
    # No conditional formatting in initial file — task-introduced change.
    # (0.10 pts)
    # -------------------------------------------------------------------------
    try:
        cf_found = False
        cf_formula_ok = False
        cf_fill_ok = False
        cf_range_ok = False

        for cf in ws.conditional_formatting:
            cf_str = str(cf)
            # Check if this CF covers A2:G9 (or contains that range)
            if 'A2:G9' in cf_str or 'A2' in cf_str:
                rules = ws.conditional_formatting[cf]
                for rule in rules:
                    if hasattr(rule, 'formula') and rule.formula:
                        formula_str = ''.join(str(f) for f in rule.formula).upper().replace(' ', '')
                        # Accept $G2>0.1 or G2>0.1 or similar
                        if 'G2>0.1' in formula_str or '$G2>0.1' in formula_str:
                            cf_formula_ok = True
                    if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                        try:
                            rgb = rule.dxf.fill.fgColor.rgb
                            # Accept FFFF6600 (opaque orange)
                            if rgb.upper() in ('FFFF6600', 'FF6600'):
                                cf_fill_ok = True
                        except Exception:
                            pass
                cf_found = True
                # Check range covers A2:G9
                if 'A2:G9' in cf_str:
                    cf_range_ok = True

        if cf_formula_ok and cf_fill_ok and cf_range_ok:
            print(f"PASS: Component 5 — Conditional formatting on A2:G9 with formula $G2>0.1 and orange fill (0.10 pts)")
            total_score += 0.10
        elif cf_formula_ok and cf_fill_ok:
            print(f"PARTIAL: Component 5 — CF formula and fill correct but range may not be exactly A2:G9 (0.05 pts)")
            total_score += 0.05
        elif cf_found:
            print(f"FAIL: Component 5 — CF found but conditions not met: formula_ok={cf_formula_ok}, fill_ok={cf_fill_ok}, range_ok={cf_range_ok}")
        else:
            print(f"FAIL: Component 5 — No conditional formatting found on A2:G9")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {round(final_score, 4)}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
