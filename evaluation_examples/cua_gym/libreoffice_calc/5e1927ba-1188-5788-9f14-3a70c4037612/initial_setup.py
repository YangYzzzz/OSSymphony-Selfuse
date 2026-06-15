"""
Initial Setup: FTE Budget Planning Model
Task ID: calc_hr_fte_budget_planning_051
Domain: libreoffice_calc

Creates a spreadsheet with an FTE budget planning sheet.
Columns B (Planned FTEs), C (Avg Salary), E (Last Year Budget) are pre-filled.
Columns D, F, G are intentionally left empty (task asks to add formulas).
Row 10 totals are intentionally absent (task asks to add them).
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_fte_budget_planning_051'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: FTE Plan ---
    ws = wb.active
    ws.title = 'FTE Plan'

    # Headers in row 1
    headers = [
        'Department',        # A
        'Planned FTEs',      # B
        'Avg Salary',        # C
        'Total Salary Cost', # D
        'Last Year Budget',  # E
        'YoY Variance',      # F
        'Variance %',        # G
    ]
    thin = Side(style='thin', color='000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Department data: 8 departments
    # Columns: Department, Planned FTEs, Avg Salary, [empty D], Last Year Budget, [empty F], [empty G]
    departments = [
        ('Engineering',          42,  118000,   None,  4750000,   None,  None),
        ('Product Management',   12,  125000,   None,  1430000,   None,  None),
        ('Sales',                55,   92000,   None,  4900000,   None,  None),
        ('Marketing',            18,   88000,   None,  1520000,   None,  None),
        ('Customer Success',     30,   78000,   None,  2200000,   None,  None),
        ('Finance & Accounting', 14,   95000,   None,  1250000,   None,  None),
        ('Human Resources',      10,   85000,   None,   820000,   None,  None),
        ('Operations',           20,   80000,   None,  1540000,   None,  None),
    ]

    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, row_data in enumerate(departments, 2):
        dept, ftes, avg_sal, total_cost, last_yr, variance, var_pct = row_data
        # Column A: Department
        cell_a = ws.cell(row=row_idx, column=1, value=dept)
        cell_a.border = cell_border
        cell_a.alignment = Alignment(horizontal='left', vertical='center')

        # Column B: Planned FTEs
        cell_b = ws.cell(row=row_idx, column=2, value=ftes)
        cell_b.border = cell_border
        cell_b.alignment = Alignment(horizontal='center', vertical='center')
        cell_b.number_format = '0'

        # Column C: Avg Salary — pre-filled
        cell_c = ws.cell(row=row_idx, column=3, value=avg_sal)
        cell_c.border = cell_border
        cell_c.alignment = Alignment(horizontal='right', vertical='center')
        cell_c.number_format = '$#,##0'

        # Column D: Total Salary Cost — intentionally EMPTY
        cell_d = ws.cell(row=row_idx, column=4, value=None)
        cell_d.border = cell_border

        # Column E: Last Year Budget — pre-filled
        cell_e = ws.cell(row=row_idx, column=5, value=last_yr)
        cell_e.border = cell_border
        cell_e.alignment = Alignment(horizontal='right', vertical='center')
        cell_e.number_format = '$#,##0'

        # Column F: YoY Variance — intentionally EMPTY
        cell_f = ws.cell(row=row_idx, column=6, value=None)
        cell_f.border = cell_border

        # Column G: Variance % — intentionally EMPTY
        cell_g = ws.cell(row=row_idx, column=7, value=None)
        cell_g.border = cell_border

    # Row 10 totals intentionally absent (task asks to add them)

    # Column widths for readability
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 14

    # Row 1 height
    ws.row_dimensions[1].height = 22

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: FTE Plan')
    print('Pre-filled: columns A (departments), B (planned FTEs), C (avg salary), E (last year budget)')
    print('Left empty: columns D (total salary cost), F (YoY variance), G (variance %)')
    print('Row 10 totals: absent')


create_initial()
