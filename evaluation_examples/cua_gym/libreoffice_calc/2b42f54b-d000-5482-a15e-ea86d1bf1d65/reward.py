"""
Reward Script: Deal Desk Approval Workflow Calculator
Task ID: calc_sales_094
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Size Level formulas in E2:E6
  Component 2 (0.25): Discount Level formulas in F2:F6
  Component 3 (0.20): Terms Level formulas in G2:G6
  Component 4 (0.15): Max Level formulas in H2:H6
  Component 5 (0.15): Approver formulas in I2:I6
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_094'


def normalize_formula(f):
    """Normalize a formula for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ""
    return f.upper().replace(" ", "")


def check_size_level_formula(formula, row):
    """
    Verify Size Level formula for given row.
    Expected: =IF(B<row><100000,1,IF(B<row><500000,2,IF(B<row><1000000,3,4)))
    """
    norm = normalize_formula(formula)
    r = row
    expected = f"=IF(B{r}<100000,1,IF(B{r}<500000,2,IF(B{r}<1000000,3,4)))"
    return norm == normalize_formula(expected)


def check_discount_level_formula(formula, row):
    """
    Verify Discount Level formula for given row.
    Expected: =IF(C<row><0.1,1,IF(C<row><0.2,2,IF(C<row><0.3,3,4)))
    """
    norm = normalize_formula(formula)
    r = row
    expected = f"=IF(C{r}<0.1,1,IF(C{r}<0.2,2,IF(C{r}<0.3,3,4)))"
    return norm == normalize_formula(expected)


def check_terms_level_formula(formula, row):
    """
    Verify Terms Level formula for given row.
    Expected: =IF(D<row>="Net 30",1,IF(D<row>="Net 60",2,3))
    """
    norm = normalize_formula(formula)
    r = row
    # The formula uses quoted strings; normalize removes spaces but keeps quotes
    expected = f'=IF(D{r}="NET30",1,IF(D{r}="NET60",2,3))'
    norm_expected = normalize_formula(expected)
    # Also accept "Net 30" with space (after normalization removes spaces)
    return norm == norm_expected


def check_max_level_formula(formula, row):
    """
    Verify Max Level formula for given row.
    Expected: =MAX(E<row>,F<row>,G<row>)
    """
    norm = normalize_formula(formula)
    r = row
    expected = f"=MAX(E{r},F{r},G{r})"
    return norm == normalize_formula(expected)


def check_approver_formula(formula, row):
    """
    Verify Approver formula for given row.
    Expected: =IF(H<row>=1,"Rep",IF(H<row>=2,"Manager",IF(H<row>=3,"Director","VP")))
    """
    norm = normalize_formula(formula)
    r = row
    expected = f'=IF(H{r}=1,"REP",IF(H{r}=2,"MANAGER",IF(H{r}=3,"DIRECTOR","VP")))'
    norm_expected = normalize_formula(expected)
    return norm == norm_expected


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check DealDesk sheet exists
    if 'DealDesk' not in wb.sheetnames:
        print("FAIL: 'DealDesk' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['DealDesk']
    data_rows = [2, 3, 4, 5, 6]  # rows with deal data

    # Component 1: Size Level formulas in E2:E6 (0.25 points)
    try:
        pass_count = 0
        for r in data_rows:
            formula = ws.cell(row=r, column=5).value  # Column E
            if formula and isinstance(formula, str) and formula.startswith('=') and check_size_level_formula(formula, r):
                pass_count += 1
            else:
                print(f"  DETAIL: E{r} expected Size Level formula, found: {repr(formula)}")
        if pass_count == 5:
            print(f"PASS: Component 1 - All 5 Size Level formulas correct (0.25 pts)")
            total_score += 0.25
        elif pass_count > 0:
            partial = round(0.25 * pass_count / 5, 3)
            print(f"PARTIAL: Component 1 - {pass_count}/5 Size Level formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - No Size Level formulas found in E2:E6")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Discount Level formulas in F2:F6 (0.25 points)
    try:
        pass_count = 0
        for r in data_rows:
            formula = ws.cell(row=r, column=6).value  # Column F
            if formula and isinstance(formula, str) and formula.startswith('=') and check_discount_level_formula(formula, r):
                pass_count += 1
            else:
                print(f"  DETAIL: F{r} expected Discount Level formula, found: {repr(formula)}")
        if pass_count == 5:
            print(f"PASS: Component 2 - All 5 Discount Level formulas correct (0.25 pts)")
            total_score += 0.25
        elif pass_count > 0:
            partial = round(0.25 * pass_count / 5, 3)
            print(f"PARTIAL: Component 2 - {pass_count}/5 Discount Level formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 - No Discount Level formulas found in F2:F6")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Terms Level formulas in G2:G6 (0.20 points)
    try:
        pass_count = 0
        for r in data_rows:
            formula = ws.cell(row=r, column=7).value  # Column G
            if formula and isinstance(formula, str) and formula.startswith('=') and check_terms_level_formula(formula, r):
                pass_count += 1
            else:
                print(f"  DETAIL: G{r} expected Terms Level formula, found: {repr(formula)}")
        if pass_count == 5:
            print(f"PASS: Component 3 - All 5 Terms Level formulas correct (0.20 pts)")
            total_score += 0.20
        elif pass_count > 0:
            partial = round(0.20 * pass_count / 5, 3)
            print(f"PARTIAL: Component 3 - {pass_count}/5 Terms Level formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 - No Terms Level formulas found in G2:G6")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Max Level formulas in H2:H6 (0.15 points)
    try:
        pass_count = 0
        for r in data_rows:
            formula = ws.cell(row=r, column=8).value  # Column H
            if formula and isinstance(formula, str) and formula.startswith('=') and check_max_level_formula(formula, r):
                pass_count += 1
            else:
                print(f"  DETAIL: H{r} expected Max Level formula, found: {repr(formula)}")
        if pass_count == 5:
            print(f"PASS: Component 4 - All 5 Max Level formulas correct (0.15 pts)")
            total_score += 0.15
        elif pass_count > 0:
            partial = round(0.15 * pass_count / 5, 3)
            print(f"PARTIAL: Component 4 - {pass_count}/5 Max Level formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 - No Max Level formulas found in H2:H6")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Approver formulas in I2:I6 (0.15 points)
    try:
        pass_count = 0
        for r in data_rows:
            formula = ws.cell(row=r, column=9).value  # Column I
            if formula and isinstance(formula, str) and formula.startswith('=') and check_approver_formula(formula, r):
                pass_count += 1
            else:
                print(f"  DETAIL: I{r} expected Approver formula, found: {repr(formula)}")
        if pass_count == 5:
            print(f"PASS: Component 5 - All 5 Approver formulas correct (0.15 pts)")
            total_score += 0.15
        elif pass_count > 0:
            partial = round(0.15 * pass_count / 5, 3)
            print(f"PARTIAL: Component 5 - {pass_count}/5 Approver formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 - No Approver formulas found in I2:I6")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook for LibreOffice
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
persist_app_state()

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
