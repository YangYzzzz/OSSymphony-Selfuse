"""
Initial Setup: Deal desk approval workflow calculator
Task ID: calc_sales_094
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_094'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: DealDesk ---
    ws = wb.active
    ws.title = 'DealDesk'

    # Headers
    headers = ['Deal', 'Value', 'Discount %', 'Payment Terms',
               'Size Level', 'Discount Level', 'Terms Level', 'Max Level', 'Approver']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Deal data (columns A-D only; E-I left empty for agent to fill)
    deals = [
        ['D1', 50000, 0.05, 'Net 30'],
        ['D2', 250000, 0.18, 'Net 60'],
        ['D3', 750000, 0.25, 'Net 90'],
        ['D4', 100000, 0.08, 'Net 30'],
        ['D5', 1500000, 0.30, 'Net 90'],
    ]

    data_align = Alignment(horizontal='center', vertical='center')

    for r, row_data in enumerate(deals, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = data_align

    # Format Value column as currency
    for r in range(2, 7):
        ws.cell(row=r, column=2).number_format = '$#,##0'

    # Format Discount % column as percentage
    for r in range(2, 7):
        ws.cell(row=r, column=3).number_format = '0%'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 14

    # Add a reference sheet with the approval rules
    ws2 = wb.create_sheet('ApprovalRules')
    ws2['A1'] = 'Deal Size Thresholds'
    ws2['A1'].font = Font(bold=True, size=12)
    rules_data = [
        ['Range', 'Level'],
        ['< $100,000', 1],
        ['$100,000 - $500,000', 2],
        ['$500,000 - $1,000,000', 3],
        ['> $1,000,000', 4],
    ]
    for r, row_data in enumerate(rules_data, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    ws2['A8'] = 'Discount Thresholds'
    ws2['A8'].font = Font(bold=True, size=12)
    disc_rules = [
        ['Range', 'Level'],
        ['< 10%', 1],
        ['10% - 20%', 2],
        ['20% - 30%', 3],
        ['> 30%', 4],
    ]
    for r, row_data in enumerate(disc_rules, 9):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    ws2['A15'] = 'Payment Terms Levels'
    ws2['A15'].font = Font(bold=True, size=12)
    terms_rules = [
        ['Terms', 'Level'],
        ['Net 30', 1],
        ['Net 60', 2],
        ['Net 90', 3],
    ]
    for r, row_data in enumerate(terms_rules, 16):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    ws2['A21'] = 'Approval Mapping'
    ws2['A21'].font = Font(bold=True, size=12)
    approval_map = [
        ['Max Level', 'Approver'],
        [1, 'Rep'],
        [2, 'Manager'],
        [3, 'Director'],
        [4, 'VP'],
    ]
    for r, row_data in enumerate(approval_map, 22):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
