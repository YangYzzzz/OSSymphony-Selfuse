"""
Reward Script: Freeze first row and first column at B2
Task ID: calc_gsi_038
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): freeze_panes is set to "B2"
  Component 2 (0.3): Pane state is "frozen" with xSplit=1 and ySplit=1
  Component 3 (0.2): Pane activePane is "bottomRight" and topLeftCell is "B2"
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_038'


def verify_task(file_path):
    """
    Verify that freeze panes are correctly set at B2, freezing both
    the first row (header) and first column (product names).
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Precondition gate: sheet should have data (not corrupted)
    if ws.max_row < 2 or ws.max_column < 2:
        print(f"CRITICAL: Sheet appears empty or corrupted (max_row={ws.max_row}, max_col={ws.max_column})")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: freeze_panes is set to "B2" (0.5 points)
    # This is the primary task requirement — freezing at B2 locks row 1 and column A.
    try:
        fp = ws.freeze_panes
        if fp is not None and str(fp) == "B2":
            print(f"PASS: Component 1 — freeze_panes is '{fp}' (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — expected freeze_panes='B2', found: '{fp}'")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Pane object has state="frozen", xSplit=1, ySplit=1 (0.3 points)
    # When freeze_panes="B2", openpyxl creates a Pane with xSplit=1 (1 col frozen)
    # and ySplit=1 (1 row frozen), state="frozen".
    try:
        sv = ws.sheet_view
        pane = sv.pane
        if pane is not None:
            x_ok = pane.xSplit == 1.0
            y_ok = pane.ySplit == 1.0
            state_ok = pane.state == "frozen"
            if x_ok and y_ok and state_ok:
                print(f"PASS: Component 2 — pane state='frozen', xSplit=1.0, ySplit=1.0 (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — pane xSplit={pane.xSplit}, ySplit={pane.ySplit}, state='{pane.state}'")
        else:
            print(f"FAIL: Component 2 — no pane object found (freeze not applied)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Pane activePane="bottomRight" and topLeftCell="B2" (0.2 points)
    # Correct B2 freeze creates 4 quadrants with bottomRight as the active editing pane,
    # and topLeftCell of the scrollable area at B2.
    try:
        sv = ws.sheet_view
        pane = sv.pane
        if pane is not None:
            active_ok = pane.activePane == "bottomRight"
            tlc_ok = str(pane.topLeftCell) == "B2"
            if active_ok and tlc_ok:
                print(f"PASS: Component 3 — activePane='bottomRight', topLeftCell='B2' (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 3 — activePane='{pane.activePane}', topLeftCell='{pane.topLeftCell}'")
        else:
            print(f"FAIL: Component 3 — no pane object found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
