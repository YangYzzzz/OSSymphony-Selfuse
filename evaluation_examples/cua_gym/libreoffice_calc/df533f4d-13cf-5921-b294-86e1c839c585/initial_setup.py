"""
Initial Setup: Freeze first row and column at B2
Task ID: calc_gsi_038
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_038'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product Matrix"

    # --- Headers in row 1 ---
    # Column A: "Product"
    # Columns B-P: Quarter labels
    quarter_headers = [
        "Q1 2023", "Q2 2023", "Q3 2023", "Q4 2023",
        "Q1 2024", "Q2 2024", "Q3 2024", "Q4 2024",
        "Q1 2025", "Q2 2025", "Q3 2025", "Q4 2025",
        "Q1 2026", "Q2 2026", "Q3 2026",
    ]
    ws.cell(row=1, column=1, value="Product")
    for col_idx, qh in enumerate(quarter_headers, 2):
        ws.cell(row=1, column=col_idx, value=qh)

    # Style headers
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    for col_idx in range(1, 17):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    # Set column widths
    ws.column_dimensions["A"].width = 28
    for col_letter in ["B","C","D","E","F","G","H","I","J","K","L","M","N","O","P"]:
        ws.column_dimensions[col_letter].width = 12

    # --- Product names (119 rows: rows 2-120) ---
    product_prefixes = [
        "ProVista", "AquaFlow", "TerraMax", "SkyLite", "NovaPrime",
        "EchoWave", "ZenCore", "BrightEdge", "FlexiGrip", "PureLink",
        "CoreSync", "VoltRise", "AeroGlide", "QuickStream", "SilverLine",
        "IronClad", "BluePeak", "RapidFire", "GreenLeaf", "OmniTech",
        "SwiftCargo", "DigiPulse", "SolarWind", "ClearPath", "TitanForce",
    ]
    product_suffixes = [
        "100", "200", "300", "Pro", "Elite",
        "X1", "X2", "S10", "S20", "Max",
        "Ultra", "Lite", "Plus", "EX", "V2",
    ]
    categories = [
        "Standard", "Premium", "Enterprise", "Compact", "Industrial",
    ]

    random.seed(42)
    products = []
    for i in range(119):
        prefix = product_prefixes[i % len(product_prefixes)]
        suffix = product_suffixes[i % len(product_suffixes)]
        cat = categories[i % len(categories)]
        products.append(f"{prefix} {suffix} ({cat})")

    # Write data rows 2-120
    for row_idx, product_name in enumerate(products, 2):
        ws.cell(row=row_idx, column=1, value=product_name)
        # Generate realistic quarterly revenue figures
        base = random.randint(5000, 150000)
        for col_idx in range(2, 17):
            variation = random.uniform(0.7, 1.3)
            trend = 1.0 + (col_idx - 2) * 0.02  # slight upward trend
            value = round(base * variation * trend, 2)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Format data cells with number format
    for row_idx in range(2, 121):
        for col_idx in range(2, 17):
            ws.cell(row=row_idx, column=col_idx).number_format = '#,##0.00'

    # NO freeze panes -- this is the task the agent must complete
    ws.freeze_panes = None

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
