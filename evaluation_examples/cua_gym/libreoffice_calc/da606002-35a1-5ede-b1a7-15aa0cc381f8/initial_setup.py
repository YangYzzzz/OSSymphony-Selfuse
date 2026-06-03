"""
Initial Setup: Protect 'Scores' sheet — create unprotected spreadsheet with student scores
Task ID: calc_ps_019
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_019'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Scores'

    # Headers
    headers = ['Name', 'Math', 'Science', 'English']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 25 students with realistic names and scores
    students = [
        ['Sarah Chen', 92, 88, 95],
        ['Marcus Johnson', 78, 85, 82],
        ['Emily Rodriguez', 95, 91, 89],
        ['James Nakamura', 67, 73, 70],
        ['Priya Patel', 88, 94, 91],
        ['David Kim', 74, 68, 77],
        ['Olivia Thompson', 91, 87, 93],
        ['Carlos Mendez', 83, 79, 85],
        ['Aisha Mohammed', 96, 92, 98],
        ['Ryan O\'Brien', 72, 65, 74],
        ['Sophie Laurent', 89, 91, 86],
        ['Wei Zhang', 94, 97, 90],
        ['Isabella Rossi', 81, 76, 83],
        ['Michael Adams', 70, 72, 68],
        ['Fatima Al-Rashid', 87, 90, 92],
        ['Tyler Washington', 76, 71, 79],
        ['Hannah Mueller', 93, 89, 94],
        ['Raj Krishnamurthy', 85, 88, 80],
        ['Grace Okafor', 79, 82, 75],
        ['Daniel Fischer', 90, 86, 88],
        ['Mei-Ling Wu', 97, 95, 96],
        ['Alexander Petrov', 73, 69, 71],
        ['Zara Hussain', 86, 93, 87],
        ['Liam O\'Connor', 80, 77, 84],
        ['Nadia Volkov', 91, 84, 90],
    ]

    for r, row_data in enumerate(students, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10

    # Sheet is NOT protected (initial state)
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
