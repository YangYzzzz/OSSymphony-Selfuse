"""
Reward Script: Protect 'Scores' sheet without password, allow sort, disallow filter
Task ID: calc_ps_019
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): Sheet protection enabled without password
  Component 2 (0.25): Sort is allowed under protection
  Component 3 (0.25): AutoFilter is NOT allowed under protection
  Component 4 (0.15): Cell editing is blocked (insertRows, insertColumns, deleteRows, deleteColumns restricted)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_019'


def persist_app_state(domain: str):
    """Attempt to save any unsaved GUI state before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            import time
            time.sleep(0.8)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    openpyxl SheetProtection attribute semantics:
    - prot.sheet: True = protection is ON
    - prot.password / prot.hashValue: None = no password
    - prot.sort: False = sort is ALLOWED (not restricted)
    - prot.autoFilter: True = autoFilter is RESTRICTED (blocked)
    - prot.insertRows: True = insert rows is RESTRICTED (blocked)
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that 'Scores' sheet exists
    if 'Scores' not in wb.sheetnames:
        print(f"FAIL: 'Scores' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Scores']
    prot = ws.protection

    # Component 1: Sheet protection enabled without password (0.35 points)
    # Initial state: protection is OFF (prot.sheet=False, no sheetProtection element)
    # Golden state: protection is ON (prot.sheet=True) with no password
    try:
        protection_on = (prot.sheet is True or prot.sheet == 1)
        no_password = (prot.password is None and prot.hashValue is None)
        if protection_on and no_password:
            print(f"PASS: Component 1 — Sheet protection ON, no password (0.35 pts)")
            total_score += 0.35
        elif protection_on and not no_password:
            print(f"FAIL: Component 1 — Protection is ON but password is set (password={prot.password!r}, hash={prot.hashValue!r})")
        else:
            print(f"FAIL: Component 1 — Protection is OFF (prot.sheet={prot.sheet})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Sort is allowed under protection (0.25 points)
    # In openpyxl, prot.sort=False means sort is NOT restricted (allowed)
    # In OOXML XML, sort="0" means sort is allowed
    # Initial state: no protection, so this component requires protection to be on
    try:
        if protection_on:
            # prot.sort: False means sort is allowed (not restricted)
            sort_allowed = (prot.sort is False or prot.sort == 0 or prot.sort is None)
            if sort_allowed:
                print(f"PASS: Component 2 — Sort is allowed (prot.sort={prot.sort}) (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 2 — Sort is restricted (prot.sort={prot.sort})")
        else:
            print(f"FAIL: Component 2 — Cannot verify sort permission: protection is OFF")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: AutoFilter is NOT allowed under protection (0.25 points)
    # In openpyxl, prot.autoFilter=True means autoFilter IS restricted (blocked)
    # In OOXML XML, autoFilter="1" means autoFilter is blocked
    # Initial state: no protection, so this component requires protection to be on
    try:
        if protection_on:
            # prot.autoFilter: True means autoFilter is restricted (blocked/not allowed)
            autofilter_blocked = (prot.autoFilter is True or prot.autoFilter == 1)
            if autofilter_blocked:
                print(f"PASS: Component 3 — AutoFilter is blocked (prot.autoFilter={prot.autoFilter}) (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 3 — AutoFilter is allowed (prot.autoFilter={prot.autoFilter})")
        else:
            print(f"FAIL: Component 3 — Cannot verify autoFilter permission: protection is OFF")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Cell editing is blocked (0.15 points)
    # When protection is on, cells are locked by default (cell.protection.locked=True)
    # We verify that general editing restrictions are in place:
    # insertRows, insertColumns, deleteRows, deleteColumns should be restricted (True)
    # AND data is still intact (25 students)
    try:
        if protection_on:
            # Check that modification operations are restricted
            insert_rows_blocked = (prot.insertRows is True or prot.insertRows == 1)
            insert_cols_blocked = (prot.insertColumns is True or prot.insertColumns == 1)
            delete_rows_blocked = (prot.deleteRows is True or prot.deleteRows == 1)
            delete_cols_blocked = (prot.deleteColumns is True or prot.deleteColumns == 1)

            blocked_count = sum([insert_rows_blocked, insert_cols_blocked,
                                delete_rows_blocked, delete_cols_blocked])

            # Also verify data integrity: 25 students in rows 2-26
            has_data = (ws.max_row >= 26 and ws.cell(row=1, column=1).value == 'Name')

            if blocked_count >= 3 and has_data:
                print(f"PASS: Component 4 — Cell editing blocked ({blocked_count}/4 operations restricted), data intact (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 4 — Only {blocked_count}/4 operations restricted, data_intact={has_data}")
        else:
            print(f"FAIL: Component 4 — Cannot verify editing restrictions: protection is OFF")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist app state before verification
persist_app_state("libreoffice_calc")

# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
