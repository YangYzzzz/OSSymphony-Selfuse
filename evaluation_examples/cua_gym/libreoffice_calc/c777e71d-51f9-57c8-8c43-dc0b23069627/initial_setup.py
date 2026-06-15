"""
Initial Setup: Content Marketing Performance Spreadsheet
Task ID: calc_sales_content_performance_062
Domain: libreoffice_calc

Creates ContentMetrics sheet with 35 content pieces.
Columns D-F populated. Columns G-I are empty (to be filled by agent).
No formulas, no conditional formatting, no sorting applied yet.
"""

import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_content_performance_062'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ContentMetrics'

    # Headers
    headers = [
        'Content Title', 'Type', 'Published Date',
        'Views', 'Leads', 'Deals Influenced',
        'View-Lead Rate', 'Lead-Deal Rate', 'Content Rank'
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 14

    # 35 content pieces - realistic blog posts, whitepapers, webinars
    # Views: 200-45000, Leads: 5-800, Deals: 1-120
    # Columns G, H, I must remain EMPTY
    data = [
        # (Content Title, Type, Published Date, Views, Leads, Deals Influenced)
        ('The Ultimate Guide to B2B Lead Generation', 'Blog Post', '2024-01-08', 42500, 780, 115),
        ('2024 State of Digital Marketing Report', 'Whitepaper', '2024-01-15', 18700, 620, 108),
        ('Mastering Content Strategy for SaaS Companies', 'Webinar', '2024-01-22', 3200, 290, 98),
        ('10 Proven Tactics for Email Marketing Success', 'Blog Post', '2024-02-05', 28300, 510, 92),
        ('Account-Based Marketing: A Comprehensive Playbook', 'Whitepaper', '2024-02-12', 12400, 440, 87),
        ('How to Build a High-Converting Landing Page', 'Blog Post', '2024-02-19', 35100, 680, 81),
        ('Sales Enablement Best Practices Webinar Series', 'Webinar', '2024-02-26', 4100, 325, 76),
        ('SEO Trends and Algorithm Updates 2024', 'Blog Post', '2024-03-04', 22800, 390, 71),
        ('Customer Journey Mapping Framework', 'Whitepaper', '2024-03-11', 9600, 310, 65),
        ('Social Media Advertising ROI Calculator Guide', 'Blog Post', '2024-03-18', 19500, 420, 60),
        ('Product-Led Growth Strategies for B2B SaaS', 'Webinar', '2024-03-25', 5800, 355, 57),
        ('Demand Generation Playbook for Enterprise Teams', 'Whitepaper', '2024-04-02', 7200, 265, 53),
        ('Video Marketing: From Strategy to Execution', 'Blog Post', '2024-04-09', 16300, 310, 49),
        ('Understanding Customer Acquisition Cost', 'Blog Post', '2024-04-16', 24700, 450, 45),
        ('Pipeline Velocity Optimization Workshop', 'Webinar', '2024-04-23', 2800, 210, 42),
        ('Marketing Attribution Models Explained', 'Whitepaper', '2024-05-01', 8900, 280, 39),
        ('Building Brand Authority Through Thought Leadership', 'Blog Post', '2024-05-08', 13600, 245, 36),
        ('Personalization at Scale: A Practical Guide', 'Whitepaper', '2024-05-15', 6700, 195, 33),
        ('Conversion Rate Optimization Techniques', 'Blog Post', '2024-05-22', 31400, 570, 30),
        ('Growth Hacking for B2B Startups Webinar', 'Webinar', '2024-05-29', 3700, 185, 28),
        ('The Power of Customer Testimonials in Sales', 'Blog Post', '2024-06-05', 11200, 198, 25),
        ('Data-Driven Marketing Decision Framework', 'Whitepaper', '2024-06-12', 5300, 162, 22),
        ('Effective Cold Outreach Email Templates', 'Blog Post', '2024-06-19', 44700, 750, 19),
        ('Marketing Automation for Mid-Market Companies', 'Webinar', '2024-06-26', 2400, 155, 17),
        ('Influencer Marketing for B2B Brands', 'Blog Post', '2024-07-03', 8700, 178, 15),
        ('Revenue Operations Alignment Strategies', 'Whitepaper', '2024-07-10', 4800, 135, 13),
        ('Community-Led Growth: Building Advocacy Programs', 'Blog Post', '2024-07-17', 7100, 142, 11),
        ('Podcast Marketing Strategy and Distribution', 'Blog Post', '2024-07-24', 5600, 98, 9),
        ('Interactive Content for Lead Qualification', 'Webinar', '2024-07-31', 1800, 112, 8),
        ('LinkedIn Advertising Best Practices 2024', 'Blog Post', '2024-08-07', 17200, 285, 7),
        ('Competitive Intelligence Gathering Methods', 'Whitepaper', '2024-08-14', 3900, 108, 6),
        ('Micro-Segmentation for Email Campaigns', 'Blog Post', '2024-08-21', 6300, 122, 5),
        ('Churn Reduction Through Customer Success', 'Webinar', '2024-08-28', 1200, 68, 3),
        ('PR and Earned Media Strategy Workshop', 'Whitepaper', '2024-09-04', 2700, 75, 2),
        ('Getting Started with Google Analytics 4', 'Blog Post', '2024-09-11', 38900, 620, 1),
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])   # Content Title
        ws.cell(row=r, column=2, value=row_data[1])   # Type
        ws.cell(row=r, column=3, value=row_data[2])   # Published Date
        ws.cell(row=r, column=4, value=row_data[3])   # Views
        ws.cell(row=r, column=5, value=row_data[4])   # Leads
        ws.cell(row=r, column=6, value=row_data[5])   # Deals Influenced
        # Columns G (7), H (8), I (9) are intentionally left EMPTY

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: ContentMetrics')
    print(f'  Rows: 1 header + 35 data rows')
    print(f'  Columns G, H, I: empty (to be filled by agent)')


create_initial()
