"""
Reward Script: Evaluate content marketing performance metrics
Task ID: calc_sales_content_performance_062
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.30): View-Lead Rate formulas in G2:G36 (=E/D pattern)
  - Component 2 (0.30): Lead-Deal Rate formulas in H2:H36 (=IF(E>0,F/E,0) pattern)
  - Component 3 (0.20): Content Rank formulas in I2:I36 (=RANK(F,fixed_range,0) pattern)
  - Component 4 (0.20): Conditional formatting highlights top 5 (rank<=5) in green across A-I
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_content_performance_062'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: ContentMetrics sheet must exist
    if 'ContentMetrics' not in wb.sheetnames:
        print("FAIL: 'ContentMetrics' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['ContentMetrics']

    # Component 1: View-Lead Rate formulas in G2:G36 (=E{row}/D{row} pattern) (0.30 points)
    # Task requires: G2:G36 = =E2/D2 (View to Lead rate, percentage format)
    # Fails on initial (all None) and passes on golden (formula strings present)
    try:
        g_formula_count = sum(
            1 for row in range(2, 37)
            if ws.cell(row=row, column=7).value
            and isinstance(ws.cell(row=row, column=7).value, str)
            and re.match(r'=E\d+/D\d+', ws.cell(row=row, column=7).value, re.IGNORECASE)
        )
        if g_formula_count == 35:
            print(f"PASS: Component 1 — View-Lead Rate: all 35 rows have =E/D formulas (0.30 pts)")
            total_score += 0.30
        elif g_formula_count >= 30:
            print(f"PARTIAL: Component 1 — View-Lead Rate: {g_formula_count}/35 rows have =E/D formulas (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — View-Lead Rate: only {g_formula_count}/35 rows have =E/D formulas")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Lead-Deal Rate formulas in H2:H36 (=IF(E>0,F/E,0) pattern) (0.30 points)
    # Task requires: H2:H36 = =IF(E2>0,F2/E2,0) (Lead to Deal rate, percentage format)
    # Fails on initial (all None) and passes on golden (formula strings present)
    try:
        h_formula_count = sum(
            1 for row in range(2, 37)
            if ws.cell(row=row, column=8).value
            and isinstance(ws.cell(row=row, column=8).value, str)
            and re.match(r'=IF\(E\d+>0,F\d+/E\d+,0\)', ws.cell(row=row, column=8).value, re.IGNORECASE)
        )
        if h_formula_count == 35:
            print(f"PASS: Component 2 — Lead-Deal Rate: all 35 rows have =IF(E>0,F/E,0) formulas (0.30 pts)")
            total_score += 0.30
        elif h_formula_count >= 30:
            print(f"PARTIAL: Component 2 — Lead-Deal Rate: {h_formula_count}/35 rows have IF formulas (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 — Lead-Deal Rate: only {h_formula_count}/35 rows have IF formulas")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Content Rank formulas in I2:I36 (=RANK(F,fixed_range,0)) (0.20 points)
    # Task requires: I2:I36 = =RANK(F2,$F$2:$F$36,0)
    # Fails on initial (all None) and passes on golden (formula strings present)
    try:
        i_formula_count = sum(
            1 for row in range(2, 37)
            if ws.cell(row=row, column=9).value
            and isinstance(ws.cell(row=row, column=9).value, str)
            and re.match(r'=RANK\(F\d+,\$F\$2:\$F\$36,0\)', ws.cell(row=row, column=9).value, re.IGNORECASE)
        )
        if i_formula_count == 35:
            print(f"PASS: Component 3 — Content Rank: all 35 rows have =RANK(F,$F$2:$F$36,0) formulas (0.20 pts)")
            total_score += 0.20
        elif i_formula_count >= 30:
            print(f"PARTIAL: Component 3 — Content Rank: {i_formula_count}/35 rows have RANK formulas (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — Content Rank: only {i_formula_count}/35 rows have RANK formulas")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Conditional formatting — top 5 by rank highlighted green (0.20 points)
    # Task requires: rows where I <= 5 highlighted green (fill) across columns A-I
    # Fails on initial (no CF rules) and passes on golden (FormulaRule with $I2<=5)
    try:
        cf_rules = ws.conditional_formatting
        # Count how many sub-checks pass
        cf_conditions_met = sum([
            # Check 1: any CF rule exists
            len(list(cf_rules)) > 0,
            # Check 2: a rule uses I<=5 formula condition
            any(
                re.search(r'\$?I\d*\s*<=\s*5', ' '.join(str(f) for f in (rule.formula or [])), re.IGNORECASE)
                for cf in cf_rules
                for rule in cf.rules
            ),
            # Check 3: the CF range covers A through I columns
            any(
                'A' in str(cf) and 'I' in str(cf)
                for cf in cf_rules
                for rule in cf.rules
                if re.search(r'\$?I\d*\s*<=\s*5', ' '.join(str(f) for f in (rule.formula or [])), re.IGNORECASE)
            ),
            # Check 4: green fill in the dxf (green channel dominant)
            any(
                (lambda rgb: rgb and int(rgb[4:6], 16) > int(rgb[2:4], 16) and int(rgb[4:6], 16) > int(rgb[6:8], 16))(
                    getattr(getattr(getattr(rule, 'dxf', None), 'fill', None), 'fgColor', None)
                    and getattr(getattr(getattr(rule, 'dxf', None), 'fill', None), 'fgColor').rgb
                )
                for cf in cf_rules
                for rule in cf.rules
                if re.search(r'\$?I\d*\s*<=\s*5', ' '.join(str(f) for f in (rule.formula or [])), re.IGNORECASE)
            ),
        ])
        if cf_conditions_met == 4:
            print(f"PASS: Component 4 — Conditional formatting: green highlight for rank<=5 rows over A:I (0.20 pts)")
            total_score += 0.20
        elif cf_conditions_met >= 2:
            print(f"PARTIAL: Component 4 — CF partially meets criteria ({cf_conditions_met}/4 checks passed) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 — Conditional formatting not correctly configured ({cf_conditions_met}/4 checks)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
