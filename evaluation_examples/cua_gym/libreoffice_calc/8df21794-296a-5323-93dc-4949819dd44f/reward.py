"""
Reward Script: Protect 'Timesheet' sheet with password 'time456', allow insert rows
Task ID: calc_ps_010
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Sheet protection is enabled
  Component 2 (0.3): Password protection is set
  Component 3 (0.3): Insert rows allowed, other structural changes prohibited
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_010'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Timesheet' sheet must exist
    if 'Timesheet' not in wb.sheetnames:
        print(f"FAIL: 'Timesheet' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Timesheet']
    prot = ws.protection

    # Component 1: Sheet protection is enabled (0.4 points)
    # In initial_env: prot.sheet == False -> FAIL
    # In golden_env: prot.sheet == True -> PASS
    try:
        if prot.sheet is True:
            print(f"PASS: Component 1 — Sheet protection is enabled (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — Sheet protection not enabled (sheet={prot.sheet})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Password is set on protection (0.3 points)
    # In initial_env: prot.password is None -> FAIL
    # In golden_env: prot.password is a hash string -> PASS
    # Note: openpyxl stores the hashed password, not the plaintext.
    # We verify that a password is present (non-None, non-empty).
    try:
        has_password = prot.password is not None and str(prot.password).strip() != ''
        if has_password:
            print(f"PASS: Component 2 — Password protection is set (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — No password set on sheet protection (password={prot.password})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Insert rows is allowed while other structural changes are prohibited (0.3 points)
    # openpyxl SheetProtection semantics when sheet=True:
    #   attribute=False means the action IS ALLOWED
    #   attribute=True means the action IS PROHIBITED
    # Task requires: insertRows allowed (False), other changes like deleteRows, insertColumns,
    # deleteColumns should be prohibited (True).
    # In initial_env: sheet=False, so protection is off -> this check should fail because
    #   we gate on sheet protection being True first.
    # In golden_env: insertRows=False (allowed), deleteRows=True (prohibited) etc.
    try:
        if prot.sheet is not True:
            # Protection not enabled, so insert rows permission is meaningless
            print(f"FAIL: Component 3 — Sheet not protected, cannot verify insert rows permission")
        else:
            insert_rows_allowed = (prot.insertRows is False or prot.insertRows == False)
            if insert_rows_allowed:
                # Check that other structural changes are still restricted
                # deleteRows, insertColumns, deleteColumns should be prohibited (True)
                other_restricted = (prot.deleteRows is True and
                                    prot.insertColumns is True and
                                    prot.deleteColumns is True)
                if other_restricted:
                    print(f"PASS: Component 3 — Insert rows allowed, other structural changes prohibited (0.3 pts)")
                    total_score += 0.3
                else:
                    # Partial: insert rows is allowed but other permissions are too permissive
                    print(f"PARTIAL: Component 3 — Insert rows allowed but other changes not fully restricted "
                          f"(deleteRows={prot.deleteRows}, insertColumns={prot.insertColumns}, "
                          f"deleteColumns={prot.deleteColumns})")
                    total_score += 0.15
            else:
                print(f"FAIL: Component 3 — Insert rows not allowed (insertRows={prot.insertRows})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
