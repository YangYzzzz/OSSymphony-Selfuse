"""
Initial Setup: Create timesheet spreadsheet (unprotected)
Task ID: calc_ps_010
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_010'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Timesheet'

    # --- Headers ---
    headers = ['Date', 'Project', 'Hours', 'Description']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Column widths ---
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 40

    # --- Timesheet data rows 2-20 ---
    data = [
        ['2025-03-03', 'Alpha Platform',     8.0, 'Backend API development for user auth module'],
        ['2025-03-04', 'Alpha Platform',     7.5, 'Database schema migration and testing'],
        ['2025-03-05', 'Beta Dashboard',     6.0, 'Frontend chart component integration'],
        ['2025-03-06', 'Beta Dashboard',     8.0, 'Dashboard layout redesign and responsive fixes'],
        ['2025-03-07', 'Client Portal',      4.5, 'Bug fixes for login flow reported by QA'],
        ['2025-03-10', 'Alpha Platform',     8.0, 'REST endpoint implementation for reports'],
        ['2025-03-11', 'Internal Tools',     3.0, 'Code review and documentation updates'],
        ['2025-03-12', 'Client Portal',      7.0, 'Payment processing integration with Stripe'],
        ['2025-03-13', 'Beta Dashboard',     8.0, 'Unit tests for data visualization module'],
        ['2025-03-14', 'Alpha Platform',     5.5, 'Performance optimization for search queries'],
        ['2025-03-17', 'Internal Tools',     6.0, 'CI/CD pipeline configuration for staging'],
        ['2025-03-18', 'Client Portal',      8.0, 'User profile settings page development'],
        ['2025-03-19', 'Beta Dashboard',     7.0, 'Accessibility audit and WCAG compliance'],
        ['2025-03-20', 'Alpha Platform',     4.0, 'Sprint planning and backlog grooming'],
        ['2025-03-21', 'Client Portal',      8.0, 'Email notification service implementation'],
        ['2025-03-24', 'Internal Tools',     5.0, 'Monitoring dashboard setup with Grafana'],
        ['2025-03-25', 'Beta Dashboard',     7.5, 'Data export feature for CSV and PDF'],
        ['2025-03-26', 'Alpha Platform',     8.0, 'API rate limiting and throttling logic'],
        ['2025-03-27', 'Client Portal',      6.5, 'End-to-end testing for checkout workflow'],
    ]

    data_font = Font(name='Calibri', size=11)
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = thin_border
            if c == 1:
                cell.number_format = 'yyyy-mm-dd'
                cell.alignment = Alignment(horizontal='center')
            elif c == 3:
                cell.number_format = '0.0'
                cell.alignment = Alignment(horizontal='center')

    # Sheet is NOT protected (task requires user to protect it)
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
