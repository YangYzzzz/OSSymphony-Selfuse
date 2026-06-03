"""
Initial Setup: Configure print settings for Inventory sheet
Task ID: calc_gg3_015
Domain: libreoffice_calc

Creates a warehouse stock spreadsheet with 300 rows of inventory data.
Print settings: portrait, no repeated rows, no page header.
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_015'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    # --- Row 1: Company name ---
    ws.merge_cells("A1:F1")
    ws["A1"] = "Westfield Distribution Co."
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # --- Row 2: Column headers ---
    headers = ["SKU", "Description", "Category", "Qty", "Location", "Value"]
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2E75B6", end_color="FF2E75B6", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[2].height = 22

    # --- Data generation for 300 rows ---
    categories = [
        "Electronics", "Hardware", "Plumbing", "Electrical", "Paint",
        "Lumber", "Automotive", "Garden", "Safety", "Cleaning",
        "HVAC", "Fasteners", "Tools", "Lighting", "Adhesives",
    ]
    locations = [
        "A-01", "A-02", "A-03", "A-04", "A-05",
        "B-01", "B-02", "B-03", "B-04", "B-05",
        "C-01", "C-02", "C-03", "C-04", "C-05",
        "D-01", "D-02", "D-03", "D-04", "D-05",
        "E-01", "E-02", "E-03", "E-04", "E-05",
    ]
    descriptions = [
        "LED Panel Light 600x600", "Copper Pipe 15mm 3m", "PVC Conduit 25mm",
        "Hex Bolt M10x50 Grade 8.8", "Silicone Sealant 300ml Clear",
        "Cordless Drill 18V Kit", "Safety Goggles Anti-Fog", "Paint Roller 9in",
        "Cable Tie 200mm Black x100", "Angle Bracket 75x75mm",
        "Wall Socket Double USB", "Pipe Wrench 14in", "Masking Tape 24mm x 50m",
        "Sandpaper 120 Grit Pack", "Wood Screws 4x40mm x200",
        "Insulation Roll 100mm", "Fire Extinguisher 2kg CO2", "Step Ladder 6ft",
        "Motion Sensor PIR", "Wire Stripper Auto",
        "Basin Mixer Tap Chrome", "Smoke Detector Ionisation", "Door Hinge 100mm Brass",
        "Fluorescent Tube T8 36W", "Duct Tape Silver 50m",
        "Air Filter HEPA 20x25", "Circuit Breaker 32A", "Cement Mix 25kg Bag",
        "Primer Undercoat 5L White", "Extension Lead 4-Gang 5m",
        "Drill Bit Set HSS 19pc", "Measuring Tape 8m", "Spirit Level 600mm",
        "Toggle Switch 10A", "Cable Clips 10mm White x100",
        "WD-40 Spray 400ml", "Hacksaw Blade 300mm", "Plasterboard 2400x1200",
        "Vinyl Gloves Large x100", "Anti-Slip Tape 50mm x 18m",
    ]

    for row_num in range(3, 303):
        sku = f"WF-{random.randint(10000, 99999)}"
        desc = random.choice(descriptions)
        cat = random.choice(categories)
        qty = random.randint(1, 500)
        loc = random.choice(locations)
        value = round(random.uniform(2.50, 450.00), 2)

        ws.cell(row=row_num, column=1, value=sku)
        ws.cell(row=row_num, column=2, value=desc)
        ws.cell(row=row_num, column=3, value=cat)
        ws.cell(row=row_num, column=4, value=qty)
        ws.cell(row=row_num, column=5, value=loc)
        val_cell = ws.cell(row=row_num, column=6, value=value)
        val_cell.number_format = '$#,##0.00'

    # --- Column widths ---
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12

    # --- Page setup: portrait, no repeated rows, no headers ---
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = False

    # Ensure NO print titles (rows to repeat) are set
    ws.print_title_rows = None

    # Ensure NO header/footer
    ws.oddHeader.center.text = ""
    ws.oddFooter.center.text = ""

    wb.save(OUTPUT)
    print(f"Initial file created: {OUTPUT}")

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print("GUI_READY: launched LibreOffice Calc with DISPLAY=:0")


create_initial()
