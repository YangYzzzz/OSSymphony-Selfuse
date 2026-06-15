"""
Reward Script: Multi-year headcount and budget projection with scenario modeling
Task ID: calc_hr_073
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25) - Scenarios sheet has formulas for HC, salary, and cost
  Component 2 (0.20) - Best Case sheet exists with correct growth rate (0.15) and salary formulas
  Component 3 (0.20) - Worst Case sheet exists with correct growth rate (0.05) and salary increase (1.05)
  Component 4 (0.20) - Summary sheet exists with cross-sheet cost references
  Component 5 (0.15) - Summary sheet has Total row with SUM formulas
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_073'


def persist_app_state(domain):
    """Save any open LibreOffice document before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            import time
            time.sleep(0.8)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def normalize_formula(f):
    """Normalize a formula string for comparison: uppercase, no spaces."""
    if not isinstance(f, str):
        return ""
    return f.upper().replace(" ", "")


def verify_task(file_path):
    """Verify task completion with progressive scoring."""
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    sheets = wb.sheetnames

    # Component 1: Scenarios sheet has formulas for HC growth, salary growth, and total cost (0.25 pts)
    try:
        comp1_score = 0.0
        if 'Scenarios' not in sheets:
            print("FAIL: Component 1 — 'Scenarios' sheet not found")
        else:
            ws = wb['Scenarios']
            # Check that B3 has a ROUND formula for HC growth
            b3 = normalize_formula(ws['B3'].value)
            if 'ROUND' in b3 and '1+' in b3.replace("1+C", "1+C"):
                comp1_score += 0.05
                print(f"PASS: Component 1a — B3 has HC growth formula: {ws['B3'].value}")
            else:
                print(f"FAIL: Component 1a — B3 expected ROUND HC formula, found: {ws['B3'].value}")

            # Check E2 has a cost formula (=B2*D2 or similar)
            e2 = normalize_formula(ws['E2'].value)
            if 'B2' in e2 and 'D2' in e2:
                comp1_score += 0.05
                print(f"PASS: Component 1b — E2 has cost formula: {ws['E2'].value}")
            else:
                print(f"FAIL: Component 1b — E2 expected cost formula (B*D), found: {ws['E2'].value}")

            # Check D3 has salary increase formula (=D2*1.03)
            d3 = normalize_formula(ws['D3'].value)
            if 'D2' in d3 and '1.03' in d3:
                comp1_score += 0.05
                print(f"PASS: Component 1c — D3 has salary increase formula: {ws['D3'].value}")
            else:
                print(f"FAIL: Component 1c — D3 expected salary formula with 1.03, found: {ws['D3'].value}")

            # Check cascading: B4 references B3, B5 references B4, etc.
            cascade_ok = 0
            for row_num in [4, 5, 6]:
                b_val = normalize_formula(ws.cell(row=row_num, column=2).value)
                prev = f"B{row_num-1}"
                if prev in b_val and 'ROUND' in b_val:
                    cascade_ok += 1
            if cascade_ok >= 2:
                comp1_score += 0.05
                print(f"PASS: Component 1d — HC formulas cascade through rows ({cascade_ok}/3)")
            else:
                print(f"FAIL: Component 1d — HC formulas don't cascade properly ({cascade_ok}/3)")

            # Check E3-E6 have cost formulas
            cost_formulas_ok = 0
            for row_num in [3, 4, 5, 6]:
                e_val = normalize_formula(ws.cell(row=row_num, column=5).value)
                b_ref = f"B{row_num}"
                d_ref = f"D{row_num}"
                if b_ref in e_val and d_ref in e_val:
                    cost_formulas_ok += 1
            if cost_formulas_ok >= 3:
                comp1_score += 0.05
                print(f"PASS: Component 1e — Cost formulas in E3:E6 ({cost_formulas_ok}/4)")
            else:
                print(f"FAIL: Component 1e — Cost formulas incomplete ({cost_formulas_ok}/4)")

        total_score += comp1_score
        print(f"Component 1 subtotal: {comp1_score}/0.25")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Best Case sheet with 15% growth rate and 3% salary increase (0.20 pts)
    try:
        comp2_score = 0.0
        # Find a sheet that represents "Best Case" scenario
        best_sheet = None
        for sn in sheets:
            if 'best' in sn.lower():
                best_sheet = sn
                break
        if best_sheet is None:
            print("FAIL: Component 2 — No 'Best Case' sheet found")
        else:
            ws_best = wb[best_sheet]
            # Check growth rate is 0.15
            c2_val = ws_best['C2'].value
            if c2_val is not None and abs(float(c2_val) - 0.15) < 0.001:
                comp2_score += 0.05
                print(f"PASS: Component 2a — Best Case growth rate is 0.15")
            else:
                print(f"FAIL: Component 2a — Best Case growth rate expected 0.15, found: {c2_val}")

            # Check HC formula exists in B3
            b3_best = normalize_formula(ws_best['B3'].value)
            if 'ROUND' in b3_best:
                comp2_score += 0.05
                print(f"PASS: Component 2b — Best Case B3 has ROUND formula")
            else:
                print(f"FAIL: Component 2b — Best Case B3 expected ROUND formula, found: {ws_best['B3'].value}")

            # Check salary formula D3 uses 1.03 (3% increase, same as base)
            d3_best = normalize_formula(ws_best['D3'].value)
            if 'D2' in d3_best and '1.03' in d3_best:
                comp2_score += 0.05
                print(f"PASS: Component 2c — Best Case D3 has 3% salary increase formula")
            else:
                print(f"FAIL: Component 2c — Best Case D3 expected 1.03 salary formula, found: {ws_best['D3'].value}")

            # Check cost formula exists in E2
            e2_best = normalize_formula(ws_best['E2'].value)
            if 'B2' in e2_best and 'D2' in e2_best:
                comp2_score += 0.05
                print(f"PASS: Component 2d — Best Case E2 has cost formula")
            else:
                print(f"FAIL: Component 2d — Best Case E2 expected cost formula, found: {ws_best['E2'].value}")

        total_score += comp2_score
        print(f"Component 2 subtotal: {comp2_score}/0.20")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Worst Case sheet with 5% growth and 5% salary increase (0.20 pts)
    try:
        comp3_score = 0.0
        worst_sheet = None
        for sn in sheets:
            if 'worst' in sn.lower():
                worst_sheet = sn
                break
        if worst_sheet is None:
            print("FAIL: Component 3 — No 'Worst Case' sheet found")
        else:
            ws_worst = wb[worst_sheet]
            # Check growth rate is 0.05
            c2_val = ws_worst['C2'].value
            if c2_val is not None and abs(float(c2_val) - 0.05) < 0.001:
                comp3_score += 0.05
                print(f"PASS: Component 3a — Worst Case growth rate is 0.05")
            else:
                print(f"FAIL: Component 3a — Worst Case growth rate expected 0.05, found: {c2_val}")

            # Check HC formula exists in B3
            b3_worst = normalize_formula(ws_worst['B3'].value)
            if 'ROUND' in b3_worst:
                comp3_score += 0.05
                print(f"PASS: Component 3b — Worst Case B3 has ROUND formula")
            else:
                print(f"FAIL: Component 3b — Worst Case B3 expected ROUND formula, found: {ws_worst['B3'].value}")

            # Check salary formula D3 uses 1.05 (5% increase for worst case)
            d3_worst = normalize_formula(ws_worst['D3'].value)
            if 'D2' in d3_worst and '1.05' in d3_worst:
                comp3_score += 0.05
                print(f"PASS: Component 3c — Worst Case D3 has 5% salary increase formula")
            else:
                print(f"FAIL: Component 3c — Worst Case D3 expected 1.05 salary formula, found: {ws_worst['D3'].value}")

            # Check cost formula exists in E2
            e2_worst = normalize_formula(ws_worst['E2'].value)
            if 'B2' in e2_worst and 'D2' in e2_worst:
                comp3_score += 0.05
                print(f"PASS: Component 3d — Worst Case E2 has cost formula")
            else:
                print(f"FAIL: Component 3d — Worst Case E2 expected cost formula, found: {ws_worst['E2'].value}")

        total_score += comp3_score
        print(f"Component 3 subtotal: {comp3_score}/0.20")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Summary sheet with cross-sheet cost references (0.20 pts)
    try:
        comp4_score = 0.0
        summary_sheet = None
        for sn in sheets:
            if 'summary' in sn.lower():
                summary_sheet = sn
                break
        if summary_sheet is None:
            print("FAIL: Component 4 — No 'Summary' sheet found")
        else:
            ws_sum = wb[summary_sheet]
            # Check that there are cross-sheet references to scenario sheets
            # B2 should reference the base case (Scenarios) sheet cost
            b2_sum = normalize_formula(ws_sum['B2'].value)
            has_base_ref = 'SCENARIOS' in b2_sum or 'BASECASE' in b2_sum.replace(" ", "")
            if has_base_ref and 'E' in b2_sum:
                comp4_score += 0.05
                print(f"PASS: Component 4a — Summary B2 references base case costs: {ws_sum['B2'].value}")
            else:
                print(f"FAIL: Component 4a — Summary B2 expected cross-sheet ref to base case, found: {ws_sum['B2'].value}")

            # C2 should reference best case sheet cost
            c2_sum = normalize_formula(ws_sum['C2'].value)
            if 'BEST' in c2_sum and 'E' in c2_sum:
                comp4_score += 0.05
                print(f"PASS: Component 4b — Summary C2 references best case costs: {ws_sum['C2'].value}")
            else:
                print(f"FAIL: Component 4b — Summary C2 expected cross-sheet ref to best case, found: {ws_sum['C2'].value}")

            # D2 should reference worst case sheet cost
            d2_sum = normalize_formula(ws_sum['D2'].value)
            if 'WORST' in d2_sum and 'E' in d2_sum:
                comp4_score += 0.05
                print(f"PASS: Component 4c — Summary D2 references worst case costs: {ws_sum['D2'].value}")
            else:
                print(f"FAIL: Component 4c — Summary D2 expected cross-sheet ref to worst case, found: {ws_sum['D2'].value}")

            # Check references exist for multiple years (at least 3 of rows 2-6)
            cross_ref_count = 0
            for row_num in [3, 4, 5, 6]:
                b_val = normalize_formula(ws_sum.cell(row=row_num, column=2).value)
                c_val = normalize_formula(ws_sum.cell(row=row_num, column=3).value)
                d_val = normalize_formula(ws_sum.cell(row=row_num, column=4).value)
                if ('SCENARIOS' in b_val or 'BASECASE' in b_val.replace(" ", "")) and 'BEST' in c_val and 'WORST' in d_val:
                    cross_ref_count += 1
            if cross_ref_count >= 3:
                comp4_score += 0.05
                print(f"PASS: Component 4d — Summary has cross-sheet refs for multiple years ({cross_ref_count}/4)")
            else:
                print(f"FAIL: Component 4d — Summary cross-sheet refs incomplete ({cross_ref_count}/4 years)")

        total_score += comp4_score
        print(f"Component 4 subtotal: {comp4_score}/0.20")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Summary sheet has Total row with SUM formulas (0.15 pts)
    try:
        comp5_score = 0.0
        if summary_sheet is None:
            print("FAIL: Component 5 — No 'Summary' sheet found")
        else:
            ws_sum = wb[summary_sheet]
            # Find a row labeled 'Total' or containing SUM formulas
            total_row = None
            for row_num in range(2, ws_sum.max_row + 1):
                a_val = ws_sum.cell(row=row_num, column=1).value
                if a_val and 'total' in str(a_val).lower():
                    total_row = row_num
                    break

            if total_row is None:
                # Also check if any row has SUM formulas across B, C, D columns
                for row_num in range(ws_sum.max_row, max(1, ws_sum.max_row - 3), -1):
                    b_val = normalize_formula(ws_sum.cell(row=row_num, column=2).value)
                    if 'SUM' in b_val:
                        total_row = row_num
                        break

            if total_row is None:
                print("FAIL: Component 5 — No Total/SUM row found in Summary sheet")
            else:
                # Check SUM formulas in B, C, D of total row
                sum_count = 0
                for col in [2, 3, 4]:
                    val = normalize_formula(ws_sum.cell(row=total_row, column=col).value)
                    if 'SUM' in val:
                        sum_count += 1
                if sum_count >= 3:
                    comp5_score = 0.15
                    print(f"PASS: Component 5 — Summary Total row {total_row} has SUM formulas in all 3 columns ({sum_count}/3)")
                elif sum_count >= 1:
                    comp5_score = 0.05
                    print(f"PARTIAL: Component 5 — Summary Total row has some SUM formulas ({sum_count}/3)")
                else:
                    print(f"FAIL: Component 5 — Total row {total_row} has no SUM formulas")

        total_score += comp5_score
        print(f"Component 5 subtotal: {comp5_score}/0.15")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
