"""
Initial Setup: Multi-year headcount and budget projection with scenario modeling
Task ID: calc_hr_073
Domain: libreoffice_calc

Creates a Scenarios sheet with base year data (2024) and empty rows for 2025-2028.
The agent must enter formulas for HC growth, salary increases, and total cost,
then create scenario copies (best/worst case) and a Summary sheet.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_073'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Scenarios ---
    ws = wb.active
    ws.title = 'Scenarios'

    # Headers
    headers = ['Year', 'Base HC', 'Growth Rate', 'Avg Salary', 'Total Cost']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Base year data (2024) - raw values, no formulas
    base_data = [2024, 150, 0.10, 85000]
    for col, val in enumerate(base_data, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Format growth rate as percentage
    ws.cell(row=2, column=3).number_format = '0%'

    # Format salary as currency
    ws.cell(row=2, column=4).number_format = '$#,##0'

    # Total Cost for 2024 is left empty (agent must enter formula)
    cell = ws.cell(row=2, column=5)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')
    cell.number_format = '$#,##0'

    # Rows for 2025-2028: only year values filled, rest empty for agent to add formulas
    for r, year in enumerate([2025, 2026, 2027, 2028], 3):
        cell = ws.cell(row=r, column=1, value=year)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        # Add borders and formatting to empty cells
        for col in range(2, 6):
            cell = ws.cell(row=r, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            if col == 3:
                cell.number_format = '0%'
            elif col in (4, 5):
                cell.number_format = '$#,##0'

    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 18

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
