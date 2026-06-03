"""
Initial Setup: Restaurant tracker spreadsheet + Chrome with Michelin Guide
Task ID: osworld_multi_apps_misc_018
Domain: libreoffice_calc (multi-app: Chrome + LibreOffice Calc)

Creates:
  - /home/user/osworld_multi_apps_misc_018.xlsx  -- visited restaurants spreadsheet
  - /home/user/michelin_guide.html               -- local mock Michelin Guide page (top 20)
Sets up GUI:
  - Chrome open with guide.michelin.com (local mock)
  - LibreOffice Calc open with restaurants.xlsx
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_misc_018'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'
HTML_PATH = f'{WORKDIR}/michelin_guide.html'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_michelin_html():
    """Create a local HTML page simulating guide.michelin.com top 20 starred restaurants."""
    html_content = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>MICHELIN Guide - New York City Starred Restaurants</title>
<style>
  body { font-family: Arial, sans-serif; margin: 20px; background: #fff; }
  h1 { color: #c00; font-size: 28px; }
  h2 { color: #333; font-size: 18px; margin-top: 30px; }
  table { width: 100%; border-collapse: collapse; margin-bottom: 20px; }
  th { background: #c00; color: white; padding: 10px; text-align: left; }
  td { padding: 8px 10px; border-bottom: 1px solid #ddd; }
  tr:nth-child(even) { background: #f9f9f9; }
  .stars { color: #c00; font-weight: bold; }
  .header-bar { background: #c00; color: white; padding: 10px 20px; margin-bottom: 20px; }
</style>
</head>
<body>
<div class="header-bar">
  <strong>MICHELIN Guide New York City 2025 - Top Starred Restaurants</strong>
</div>
<h1>Top 20 MICHELIN Starred Restaurants - New York City</h1>
<p>The MICHELIN Guide's top-rated restaurants for culinary excellence in New York City.</p>

<h2>&#9733;&#9733;&#9733; Three MICHELIN Stars (Exceptional cuisine, worth a special journey)</h2>
<table>
  <thead>
    <tr><th>#</th><th>Name</th><th>Cuisine</th><th>Stars</th><th>Location</th></tr>
  </thead>
  <tbody>
    <tr><td>1</td><td>Le Bernardin</td><td>French Seafood</td><td class="stars">3</td><td>Midtown</td></tr>
    <tr><td>2</td><td>Eleven Madison Park</td><td>Contemporary American</td><td class="stars">3</td><td>Flatiron</td></tr>
    <tr><td>3</td><td>Per Se</td><td>Contemporary American</td><td class="stars">3</td><td>Columbus Circle</td></tr>
  </tbody>
</table>

<h2>&#9733;&#9733; Two MICHELIN Stars (Excellent cuisine, worth a detour)</h2>
<table>
  <thead>
    <tr><th>#</th><th>Name</th><th>Cuisine</th><th>Stars</th><th>Location</th></tr>
  </thead>
  <tbody>
    <tr><td>4</td><td>The Modern</td><td>Contemporary American</td><td class="stars">2</td><td>Midtown</td></tr>
    <tr><td>5</td><td>Aquavit</td><td>Scandinavian</td><td class="stars">2</td><td>Midtown</td></tr>
    <tr><td>6</td><td>Marea</td><td>Italian Seafood</td><td class="stars">2</td><td>Upper West Side</td></tr>
    <tr><td>7</td><td>Gabriel Kreuther</td><td>Alsatian</td><td class="stars">2</td><td>Midtown</td></tr>
    <tr><td>8</td><td>Atomix</td><td>Korean</td><td class="stars">2</td><td>Gramercy</td></tr>
  </tbody>
</table>

<h2>&#9733; One MICHELIN Star (A very good restaurant in its category)</h2>
<table>
  <thead>
    <tr><th>#</th><th>Name</th><th>Cuisine</th><th>Stars</th><th>Location</th></tr>
  </thead>
  <tbody>
    <tr><td>9</td><td>Gramercy Tavern</td><td>American</td><td class="stars">1</td><td>Flatiron</td></tr>
    <tr><td>10</td><td>The NoMad</td><td>Contemporary American</td><td class="stars">1</td><td>NoMad</td></tr>
    <tr><td>11</td><td>Atera</td><td>New Nordic</td><td class="stars">1</td><td>Tribeca</td></tr>
    <tr><td>12</td><td>Blue Hill</td><td>Farm-to-Table</td><td class="stars">1</td><td>Greenwich Village</td></tr>
    <tr><td>13</td><td>Craft</td><td>American</td><td class="stars">1</td><td>Flatiron</td></tr>
    <tr><td>14</td><td>Daniel</td><td>French</td><td class="stars">1</td><td>Upper East Side</td></tr>
    <tr><td>15</td><td>Del Posto</td><td>Italian</td><td class="stars">1</td><td>Chelsea</td></tr>
    <tr><td>16</td><td>Gotham Bar and Grill</td><td>New American</td><td class="stars">1</td><td>Greenwich Village</td></tr>
    <tr><td>17</td><td>Jean-Georges</td><td>French</td><td class="stars">1</td><td>Columbus Circle</td></tr>
    <tr><td>18</td><td>Keens Steakhouse</td><td>American Steakhouse</td><td class="stars">1</td><td>Midtown</td></tr>
    <tr><td>19</td><td>L'Artusi</td><td>Italian</td><td class="stars">1</td><td>West Village</td></tr>
    <tr><td>20</td><td>Momofuku Ko</td><td>Contemporary American</td><td class="stars">1</td><td>East Village</td></tr>
  </tbody>
</table>

<p><em>Data from MICHELIN Guide 2025. Rankings represent editorial assessment.</em></p>
</body>
</html>
"""
    with open(HTML_PATH, 'w') as f:
        f.write(html_content)
    print(f'Michelin Guide HTML created: {HTML_PATH}')


def create_initial():
    """Create the visited restaurants spreadsheet."""
    wb = openpyxl.Workbook()

    # --- Sheet: Visited Restaurants ---
    ws = wb.active
    ws.title = 'restaurants'

    # Column headers
    headers = ['Name', 'Cuisine', 'Stars', 'Location']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Visited restaurants data (mix of Michelin and non-Michelin places)
    # From the Michelin Top 20, these are already visited:
    #   Le Bernardin (3), The Modern (2), Gabriel Kreuther (2),
    #   Gramercy Tavern (1), Daniel (1), Keens Steakhouse (1)
    # Plus some non-Michelin places to make the list realistic
    visited_data = [
        # Michelin-starred (already visited — should NOT appear in bucket_list)
        ('Le Bernardin', 'French Seafood', 3, 'Midtown'),
        ('The Modern', 'Contemporary American', 2, 'Midtown'),
        ('Gabriel Kreuther', 'Alsatian', 2, 'Midtown'),
        ('Gramercy Tavern', 'American', 1, 'Flatiron'),
        ('Daniel', 'French', 1, 'Upper East Side'),
        ('Keens Steakhouse', 'American Steakhouse', 1, 'Midtown'),
        # Non-Michelin places visited (not in top 20, so irrelevant to the task)
        ("Joe's Pizza", 'Italian Pizza', 0, 'Greenwich Village'),
        ("Katz's Delicatessen", 'American Deli', 0, 'Lower East Side'),
        ('Peter Luger Steak House', 'American Steakhouse', 0, 'Williamsburg'),
        ('Shake Shack', 'American Burgers', 0, 'Flatiron'),
        ('Momofuku Noodle Bar', 'Japanese Noodle', 0, 'East Village'),
        ('The Spotted Pig', 'British Gastropub', 0, 'West Village'),
        ('Balthazar', 'French Brasserie', 0, 'SoHo'),
        ('Carbone', 'Italian American', 0, 'Greenwich Village'),
        ("Di Fara Pizza", 'Italian Pizza', 0, 'Midwood, Brooklyn'),
    ]

    for r, row_data in enumerate(visited_data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    wb.save(OUTPUT)
    print(f'Initial restaurants file created: {OUTPUT}')

    # Kill any existing LibreOffice and Chrome instances first for clean start
    subprocess.run(['pkill', '-f', 'soffice'], capture_output=True)
    time.sleep(1)

    # Launch Chrome with the local Michelin Guide HTML file
    # Use file:// protocol to open local HTML
    launch_gui(f'google-chrome "file://{HTML_PATH}"', delay_sec=3.0)

    # Launch LibreOffice Calc with the restaurants file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)

    print('GUI_READY: launched Chrome (Michelin Guide) and LibreOffice Calc with DISPLAY=:0')


create_michelin_html()
create_initial()
