"""
Reward Script: Add Michelin-starred restaurants not yet visited to a 'bucket_list' sheet
Task ID: osworld_multi_apps_misc_018
Domain: libreoffice_calc
Scoring:
  - Component 1: 'bucket_list' sheet exists in restaurants.xlsx              (0.30 pts)
  - Component 2: Correct headers (Name, Cuisine, Stars, Location) in row 1   (0.20 pts)
  - Component 3: All 14 expected restaurants present with correct data        (0.30 pts)
  - Component 4: Data rows are sorted by Stars descending                     (0.20 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_misc_018'

# The 14 expected restaurants in the bucket_list (not already visited),
# derived from Michelin Guide top 20 minus restaurants already in the 'restaurants' sheet.
# Sorted by Stars descending, then by Michelin rank order for same-star groups.
EXPECTED_RESTAURANTS = [
    ('Eleven Madison Park', 'Contemporary American', 3, 'Flatiron'),
    ('Per Se',              'Contemporary American', 3, 'Columbus Circle'),
    ('Aquavit',             'Scandinavian',          2, 'Midtown'),
    ('Marea',               'Italian Seafood',       2, 'Upper West Side'),
    ('Atomix',              'Korean',                2, 'Gramercy'),
    ('The NoMad',           'Contemporary American', 1, 'NoMad'),
    ('Atera',               'New Nordic',            1, 'Tribeca'),
    ('Blue Hill',           'Farm-to-Table',         1, 'Greenwich Village'),
    ('Craft',               'American',              1, 'Flatiron'),
    ('Del Posto',           'Italian',               1, 'Chelsea'),
    ('Gotham Bar and Grill','New American',           1, 'Greenwich Village'),
    ('Jean-Georges',        'French',                1, 'Columbus Circle'),
    ("L'Artusi",            'Italian',               1, 'West Village'),
    ('Momofuku Ko',         'Contemporary American', 1, 'East Village'),
]

EXPECTED_HEADERS = ('Name', 'Cuisine', 'Stars', 'Location')
EXPECTED_NAMES_SET = {r[0] for r in EXPECTED_RESTAURANTS}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # --- Load workbook ---
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: 'bucket_list' sheet exists (0.30 points)
    # This is the primary task-introduced change — adding a new sheet.
    try:
        if 'bucket_list' in wb.sheetnames:
            print("PASS: Component 1 — 'bucket_list' sheet exists (0.30 pts)")
            total_score += 0.30
            ws_bucket = wb['bucket_list']
        else:
            print(f"FAIL: Component 1 — 'bucket_list' sheet not found. Sheets: {wb.sheetnames}")
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Component 2: Headers in row 1 of bucket_list are correct (0.20 points)
    try:
        headers = tuple(ws_bucket.cell(row=1, column=c).value for c in range(1, 5))
        if headers == EXPECTED_HEADERS:
            print(f"PASS: Component 2 — Headers correct: {headers} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 — Headers mismatch. Expected {EXPECTED_HEADERS}, found {headers}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: All 14 expected restaurants are present with correct data (0.30 points)
    # We read data rows from row 2 onward, stopping at first empty Name.
    try:
        data_rows = []
        for row in ws_bucket.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                break
            data_rows.append(row)

        # Build a dict of name -> (cuisine, stars, location) from the sheet
        actual_dict = {}
        for row in data_rows:
            name, cuisine, stars, location = row[0], row[1], row[2], row[3]
            if name is not None:
                actual_dict[str(name)] = (str(cuisine) if cuisine else None,
                                          int(stars) if stars is not None else None,
                                          str(location) if location else None)

        # Check all expected restaurants are present with correct data
        correct_count = 0
        missing = []
        wrong_data = []
        for exp_name, exp_cuisine, exp_stars, exp_location in EXPECTED_RESTAURANTS:
            if exp_name not in actual_dict:
                missing.append(exp_name)
            else:
                act_cuisine, act_stars, act_location = actual_dict[exp_name]
                if (act_cuisine == exp_cuisine and
                        act_stars == exp_stars and
                        act_location == exp_location):
                    correct_count += 1
                else:
                    wrong_data.append(
                        f"{exp_name}: expected ({exp_cuisine},{exp_stars},{exp_location}), "
                        f"found ({act_cuisine},{act_stars},{act_location})"
                    )

        # Also check no extra restaurants appear (visited ones should not be listed)
        extra = [n for n in actual_dict if n not in EXPECTED_NAMES_SET]

        if correct_count == len(EXPECTED_RESTAURANTS) and not extra:
            print(f"PASS: Component 3 — All {len(EXPECTED_RESTAURANTS)} expected restaurants present "
                  f"with correct data (0.30 pts)")
            total_score += 0.30
        else:
            if missing:
                print(f"FAIL: Component 3 — Missing restaurants: {missing}")
            if wrong_data:
                print(f"FAIL: Component 3 — Wrong data: {wrong_data}")
            if extra:
                print(f"FAIL: Component 3 — Unexpected (already-visited) restaurants in bucket_list: {extra}")
            if correct_count > 0:
                print(f"INFO: Component 3 — {correct_count}/{len(EXPECTED_RESTAURANTS)} restaurants correct")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Data rows are sorted by Stars descending (0.20 points)
    # Rows with higher Stars must come before rows with lower Stars.
    try:
        star_values = []
        for row in ws_bucket.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                break
            stars = row[2]
            if stars is not None:
                star_values.append(int(stars))

        if len(star_values) == 0:
            print("FAIL: Component 4 — No data rows found to check sort order")
        elif star_values == sorted(star_values, reverse=True):
            print(f"PASS: Component 4 — Rows sorted by Stars descending: {star_values} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 4 — Stars not sorted descending. Actual order: {star_values}")
            print(f"      Expected order: {sorted(star_values, reverse=True)}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 2)}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the VM environment
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
