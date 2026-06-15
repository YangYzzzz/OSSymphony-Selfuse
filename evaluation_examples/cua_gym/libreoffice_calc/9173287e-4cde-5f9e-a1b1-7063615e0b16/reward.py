"""
Reward Script: Design an invoice template with merged company header, item table,
tax calculation, and print-ready borders.
Task ID: calc_gpm_035
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Merged header cells (A1:F1, A2:F2, A3:F3) with proper formatting
  Component 2 (0.15): Header row 11 styling (bold, dark gray fill, white text, centered, borders)
  Component 3 (0.20): Amount formulas in E12:E16 (=C*D) and currency format on D/E columns
  Component 4 (0.20): Subtotal/Tax/Total formulas (E18, E19, E20) and TOTAL DUE styling
  Component 5 (0.10): Print area set to A1:F22
  Component 6 (0.10): Payment terms row 22 (italic, 9pt, centered, merged)
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_035'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Invoice' sheet must exist
    if 'Invoice' not in wb.sheetnames:
        print("FAIL: Sheet 'Invoice' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Invoice']

    # Helper: check if a given range is merged
    def is_merged(range_str):
        for mr in ws.merged_cells.ranges:
            if str(mr) == range_str:
                return True
        return False

    # =========================================================================
    # Component 1: Merged header cells with proper formatting (0.25 points)
    # Checks: A1:F1 merged + 18pt bold centered, A2:F2 merged + 10pt centered,
    #         A3:F3 merged + 10pt centered + thick bottom border
    # Initial state: no merges, no special font sizes, no borders
    # =========================================================================
    try:
        comp1_score = 0.0

        # Check A1:F1 merge + formatting
        if is_merged('A1:F1'):
            a1 = ws['A1']
            if a1.font.bold and a1.font.size and a1.font.size >= 17:
                comp1_score += 0.08
                print(f"PASS: A1:F1 merged, bold, size={a1.font.size}")
            else:
                print(f"FAIL: A1:F1 merged but formatting wrong: bold={a1.font.bold}, size={a1.font.size}")
        else:
            print("FAIL: A1:F1 not merged")

        # Check A2:F2 merge + formatting
        if is_merged('A2:F2'):
            a2 = ws['A2']
            if a2.font.size and a2.font.size <= 11:
                comp1_score += 0.06
                print(f"PASS: A2:F2 merged, size={a2.font.size}")
            else:
                print(f"FAIL: A2:F2 merged but size wrong: {a2.font.size}")
        else:
            print("FAIL: A2:F2 not merged")

        # Check A3:F3 merge + formatting
        if is_merged('A3:F3'):
            a3 = ws['A3']
            if a3.font.size and a3.font.size <= 11:
                comp1_score += 0.06
                print(f"PASS: A3:F3 merged, size={a3.font.size}")
            else:
                print(f"FAIL: A3:F3 merged but size wrong: {a3.font.size}")
        else:
            print("FAIL: A3:F3 not merged")

        # Check thick bottom border on row 3
        a3_cell = ws['A3']
        if not isinstance(a3_cell, MergedCell) and a3_cell.border.bottom.style in ('thick', 'medium'):
            comp1_score += 0.05
            print(f"PASS: Row 3 has thick bottom border: {a3_cell.border.bottom.style}")
        else:
            print(f"FAIL: Row 3 bottom border not thick: {a3_cell.border.bottom.style if not isinstance(a3_cell, MergedCell) else 'merged cell'}")

        if comp1_score > 0:
            total_score += comp1_score
        print(f"Component 1 subtotal: {comp1_score}/0.25")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Header row 11 styling (0.15 points)
    # Checks: bold, dark gray fill (FF404040), white text, centered, borders
    # Initial state: plain text, no styling
    # =========================================================================
    try:
        comp2_score = 0.0
        header_checks_passed = 0

        for col in range(1, 6):  # A11 to E11
            cell = ws.cell(row=11, column=col)
            is_bold = cell.font.bold
            is_centered = (cell.alignment.horizontal == 'center')

            # Check fill color (dark gray ~ FF404040)
            try:
                fg = cell.fill.fgColor.rgb
                has_fill = (fg is not None and fg != '00000000')
            except:
                has_fill = False

            # Check font color (white)
            try:
                fc = cell.font.color.rgb
                has_white_text = (fc is not None and 'FFFFFF' in str(fc).upper())
            except:
                has_white_text = False

            # Check border
            has_border = cell.border.bottom.style is not None

            if is_bold and is_centered and has_fill and has_white_text and has_border:
                header_checks_passed += 1

        if header_checks_passed >= 4:
            comp2_score = 0.15
            print(f"PASS: Header row 11 fully styled ({header_checks_passed}/5 headers correct)")
        elif header_checks_passed >= 2:
            comp2_score = 0.08
            print(f"PARTIAL: Header row 11 partially styled ({header_checks_passed}/5 headers correct)")
        else:
            print(f"FAIL: Header row 11 not properly styled ({header_checks_passed}/5 headers correct)")

        if comp2_score > 0:
            total_score += comp2_score
        print(f"Component 2 subtotal: {comp2_score}/0.15")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Amount formulas in E12:E16 and currency format (0.20 points)
    # Checks: E12-E16 contain =C*D formulas, D and E columns have $#,##0.00 format
    # Initial state: no formulas in E column, no currency format on D/E
    # =========================================================================
    try:
        comp3_score = 0.0

        # Check E12:E16 formulas
        formula_count = 0
        for row in range(12, 17):
            val = ws.cell(row=row, column=5).value  # column E
            if val and isinstance(val, str) and '=' in val:
                # Check it references C and D columns of same row
                val_upper = val.upper().replace(' ', '')
                expected = f'=C{row}*D{row}'
                if val_upper == expected.upper():
                    formula_count += 1

        if formula_count >= 5:
            comp3_score += 0.12
            print(f"PASS: All 5 amount formulas correct in E12:E16")
        elif formula_count >= 3:
            comp3_score += 0.06
            print(f"PARTIAL: {formula_count}/5 amount formulas correct")
        else:
            print(f"FAIL: Only {formula_count}/5 amount formulas found")

        # Check currency format on D/E columns (at least in data rows)
        currency_cells = 0
        for row in range(12, 17):
            for col in [4, 5]:  # D, E
                nf = ws.cell(row=row, column=col).number_format
                if nf and '$' in str(nf):
                    currency_cells += 1

        if currency_cells >= 8:
            comp3_score += 0.08
            print(f"PASS: Currency format applied ({currency_cells}/10 cells)")
        elif currency_cells >= 4:
            comp3_score += 0.04
            print(f"PARTIAL: Currency format partially applied ({currency_cells}/10 cells)")
        else:
            print(f"FAIL: Currency format missing ({currency_cells}/10 cells)")

        if comp3_score > 0:
            total_score += comp3_score
        print(f"Component 3 subtotal: {comp3_score}/0.20")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: Subtotal/Tax/Total formulas and TOTAL DUE styling (0.20 points)
    # Checks: E18=SUM(E12:E16), E19=E18*0.085, E20=E18+E19,
    #         A20 bold 12pt, E20 bold 12pt double underline
    # Initial state: rows 18-20 do not exist
    # =========================================================================
    try:
        comp4_score = 0.0

        # Check E18 formula (Subtotal)
        e18 = ws['E18'].value
        if e18 and isinstance(e18, str) and 'SUM' in e18.upper() and 'E12' in e18.upper() and 'E16' in e18.upper():
            comp4_score += 0.05
            print(f"PASS: E18 subtotal formula: {e18}")
        else:
            print(f"FAIL: E18 expected SUM formula, found: {e18}")

        # Check E19 formula (Tax)
        e19 = ws['E19'].value
        if e19 and isinstance(e19, str) and 'E18' in e19.upper() and '0.085' in e19:
            comp4_score += 0.05
            print(f"PASS: E19 tax formula: {e19}")
        else:
            print(f"FAIL: E19 expected tax formula, found: {e19}")

        # Check E20 formula (Total)
        e20 = ws['E20'].value
        if e20 and isinstance(e20, str) and 'E18' in e20.upper() and 'E19' in e20.upper():
            comp4_score += 0.04
            print(f"PASS: E20 total formula: {e20}")
        else:
            print(f"FAIL: E20 expected total formula, found: {e20}")

        # Check TOTAL DUE styling: A20 bold 12pt, E20 bold 12pt double underline
        a20_cell = ws['A20']
        e20_cell = ws['E20']

        # A20 may be the top-left of a merge
        if not isinstance(a20_cell, MergedCell):
            if a20_cell.font.bold and a20_cell.font.size and a20_cell.font.size >= 12:
                comp4_score += 0.03
                print(f"PASS: A20 bold 12pt")
            else:
                print(f"FAIL: A20 styling: bold={a20_cell.font.bold}, size={a20_cell.font.size}")
        else:
            print(f"INFO: A20 is a MergedCell, skipping style check")

        if not isinstance(e20_cell, MergedCell):
            if e20_cell.font.bold and e20_cell.font.underline == 'double':
                comp4_score += 0.03
                print(f"PASS: E20 bold + double underline")
            else:
                print(f"FAIL: E20 styling: bold={e20_cell.font.bold}, underline={e20_cell.font.underline}")
        else:
            print(f"INFO: E20 is a MergedCell, skipping style check")

        if comp4_score > 0:
            total_score += comp4_score
        print(f"Component 4 subtotal: {comp4_score}/0.20")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # =========================================================================
    # Component 5: Print area set to A1:F22 (0.10 points)
    # Initial state: no print area
    # =========================================================================
    try:
        comp5_score = 0.0
        pa = ws.print_area
        if pa:
            pa_str = str(pa).upper()
            # Normalize: could be "'Invoice'!$A$1:$F$22" or "$A$1:$F$22" or "A1:F22"
            if 'A' in pa_str and 'F' in pa_str and '1' in pa_str and '22' in pa_str:
                comp5_score = 0.10
                print(f"PASS: Print area set: {pa}")
            else:
                print(f"FAIL: Print area exists but unexpected: {pa}")
        else:
            print("FAIL: No print area set")

        if comp5_score > 0:
            total_score += comp5_score
        print(f"Component 5 subtotal: {comp5_score}/0.10")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # =========================================================================
    # Component 6: Payment terms row 22 (0.10 points)
    # Checks: A22:F22 merged, italic, 9pt, centered, contains payment text
    # Initial state: row 22 does not exist
    # =========================================================================
    try:
        comp6_score = 0.0

        a22 = ws['A22']
        if a22.value and 'Payment' in str(a22.value):
            # Check merge
            if is_merged('A22:F22'):
                comp6_score += 0.04
                print("PASS: A22:F22 merged with payment text")
            else:
                print("FAIL: A22:F22 not merged")

            # Check italic + small font
            if not isinstance(a22, MergedCell):
                if a22.font.italic:
                    comp6_score += 0.03
                    print(f"PASS: A22 italic")
                else:
                    print(f"FAIL: A22 not italic")

                if a22.font.size and a22.font.size <= 10:
                    comp6_score += 0.03
                    print(f"PASS: A22 size={a22.font.size}")
                else:
                    print(f"FAIL: A22 size={a22.font.size}")
            else:
                print("INFO: A22 is MergedCell, cannot check style")
        else:
            print(f"FAIL: A22 does not contain payment terms text: {a22.value}")

        if comp6_score > 0:
            total_score += comp6_score
        print(f"Component 6 subtotal: {comp6_score}/0.10")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved LibreOffice state
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
