"""
Initial Setup: Design an invoice template with merged header, item table, tax, and borders.
Task ID: calc_gpm_035
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_035'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # Provide raw company and client data in a simple unformatted layout
    # The agent's task is to turn this into a professional invoice template

    # Company info (unformatted, not merged)
    ws["A1"] = "CREATIVE SOLUTIONS LLC"
    ws["A2"] = "123 Design Avenue, Suite 400, Portland, OR 97201"
    ws["A3"] = "Phone: (503) 555-0142 | Email: billing@creativesolutions.com"

    # Invoice details (unformatted)
    ws["A5"] = "Invoice #:"
    ws["B5"] = "INV-2026-0047"
    ws["D5"] = "Date:"
    ws["E5"] = "2026-04-01"
    ws["A6"] = "Bill To:"
    ws["D6"] = "Payment Due:"
    ws["A7"] = "Acme Corporation"
    ws["D7"] = "2026-05-01"
    ws["A8"] = "456 Business Blvd"
    ws["A9"] = "Chicago, IL 60601"

    # Item table headers (unformatted - no bold, no fill, no borders, no centering)
    ws["A11"] = "Item #"
    ws["B11"] = "Description"
    ws["C11"] = "Qty"
    ws["D11"] = "Rate"
    ws["E11"] = "Amount"

    # Item data rows (raw values, no formulas in Amount column)
    items = [
        [1, "Brand Strategy Workshop", 2, 1500, ""],
        [2, "Logo Design Package", 1, 3500, ""],
        [3, "Website Wireframes", 5, 800, ""],
        [4, "Social Media Kit", 1, 2200, ""],
        [5, "Print Collateral Design", 3, 650, ""],
    ]
    for r, row_data in enumerate(items, 12):
        for c, val in enumerate(row_data, 1):
            if val != "":
                ws.cell(row=r, column=c, value=val)

    # No subtotal/tax/total rows
    # No merged cells
    # No formatting, borders, or number formats
    # No print area set
    # No payment terms footer

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
