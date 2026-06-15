"""
Reward Script: Wedding Event Budget & Vendor Tracker
Task ID: calc_grs_029
Domain: libreoffice_calc
Scoring:
  Component 1: Two sheets exist with correct names (0.10)
  Component 2: 8 category sections with vendor data and subtotals (0.25)
  Component 3: Correct column structure and header row (0.10)
  Component 4: Grand total formulas and budget tracking (0.20)
  Component 5: Conditional formatting on Budget Tracker sheet (0.10)
  Component 6: Payment Timeline sheet with sorted vendor entries (0.15)
  Component 7: Conditional formatting on Payment Timeline (0.10)
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_029'


def persist_app_state(domain):
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print(f"PERSIST: ctrl+s sent for {domain}")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    sheet_names = wb.sheetnames

    # Component 1: Two sheets exist — "Budget Tracker" and "Payment Timeline" (0.10 points)
    # Initial file has only "Sheet1", so this checks task-introduced changes.
    try:
        has_budget = any("budget" in s.lower() for s in sheet_names)
        has_timeline = any("timeline" in s.lower() or "payment" in s.lower() for s in sheet_names)
        has_two_sheets = len(sheet_names) >= 2

        if has_budget and has_timeline and has_two_sheets:
            print(f"PASS: Component 1 — Found budget and timeline sheets in {sheet_names} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 — Expected 'Budget Tracker' and 'Payment Timeline' sheets, found: {sheet_names}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Find the budget tracker sheet (flexible name matching)
    budget_ws = None
    timeline_ws = None
    for sn in sheet_names:
        if "budget" in sn.lower() or "tracker" in sn.lower():
            budget_ws = wb[sn]
        if "timeline" in sn.lower() or "payment" in sn.lower():
            timeline_ws = wb[sn]

    if budget_ws is None:
        # Fallback: try first sheet if it has substantial data
        if len(sheet_names) >= 1:
            ws0 = wb[sheet_names[0]]
            if ws0.max_row >= 20:
                budget_ws = ws0
        if budget_ws is None:
            print("FAIL: No budget tracker sheet found. Remaining checks skipped.")
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score

    # Component 2: 8 category sections with vendor data and subtotals (0.25 points)
    # Task requires: Venue, Catering, Photography/Video, Flowers/Decor,
    # Entertainment, Invitations, Transportation, Miscellaneous
    try:
        required_categories = [
            "venue", "catering", "photography", "flower", "decor",
            "entertainment", "invitation", "transportation", "miscellaneous"
        ]
        # Scan column A for category headers
        found_categories = set()
        subtotal_count = 0
        vendor_count = 0

        for row in range(1, budget_ws.max_row + 1):
            cell_val = budget_ws.cell(row=row, column=1).value
            if cell_val is None:
                continue
            cell_str = str(cell_val).lower().strip()

            # Check for category names
            for cat in required_categories:
                if cat in cell_str and "subtotal" not in cell_str and "total" not in cell_str:
                    found_categories.add(cat)

            # Count subtotal rows (they have SUM formulas in column C)
            c_val = budget_ws.cell(row=row, column=3).value
            if c_val and isinstance(c_val, str) and "SUM" in c_val.upper() and "subtotal" in cell_str:
                subtotal_count += 1

            # Count vendor rows (have a quoted price as number in column C, not a formula)
            if c_val and isinstance(c_val, (int, float)) and not isinstance(c_val, bool):
                vendor_count += 1

        # Need at least 6 of the required categories (flower/decor may be combined)
        cat_score = min(len(found_categories) / 7.0, 1.0)
        # Need at least 5 subtotals
        subtotal_score = min(subtotal_count / 6.0, 1.0)
        # Need at least 10 vendors (task says 15 vendors)
        vendor_score = min(vendor_count / 10.0, 1.0)

        comp2_score = (cat_score * 0.10) + (subtotal_score * 0.08) + (vendor_score * 0.07)
        if comp2_score > 0:
            print(f"PASS: Component 2 — Found {len(found_categories)} categories, "
                  f"{subtotal_count} subtotals, {vendor_count} vendors ({comp2_score:.2f} pts)")
            total_score += comp2_score
        else:
            print(f"FAIL: Component 2 — categories={len(found_categories)}, "
                  f"subtotals={subtotal_count}, vendors={vendor_count}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Correct column structure with headers (0.10 points)
    # Task requires columns: Vendor Name, Service Description, Quoted Price,
    # Deposit Paid, Balance Due, Payment Due Date, Contract Signed, Notes, Budget vs Actual
    try:
        # Find header row (look for a row containing "Vendor" and "Quoted" or similar)
        header_row = None
        for row in range(1, min(budget_ws.max_row + 1, 10)):
            vals = []
            for col in range(1, budget_ws.max_column + 1):
                v = budget_ws.cell(row=row, column=col).value
                if v:
                    vals.append(str(v).lower())
            combined = " ".join(vals)
            if "vendor" in combined and ("price" in combined or "quoted" in combined):
                header_row = row
                break

        if header_row:
            headers = []
            for col in range(1, budget_ws.max_column + 1):
                v = budget_ws.cell(row=header_row, column=col).value
                if v:
                    headers.append(str(v).lower())

            combined_headers = " ".join(headers)
            required_keywords = ["vendor", "description", "price", "deposit", "balance", "due", "contract", "notes"]
            found_keywords = sum(1 for kw in required_keywords if kw in combined_headers)

            # Check if header row is bold
            header_bold = budget_ws.cell(row=header_row, column=1).font.bold

            kw_score = min(found_keywords / 6.0, 1.0) * 0.07
            bold_score = 0.03 if header_bold else 0.0

            comp3_score = kw_score + bold_score
            if comp3_score > 0:
                print(f"PASS: Component 3 — Found {found_keywords}/8 header keywords, "
                      f"bold={header_bold} ({comp3_score:.2f} pts)")
                total_score += comp3_score
            else:
                print(f"FAIL: Component 3 — Only {found_keywords} header keywords found")
        else:
            print(f"FAIL: Component 3 — No header row found with 'Vendor' and 'Price'")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Grand total formulas and budget tracking (0.20 points)
    # Task requires: grand total, budget of $25,000, budget vs actual comparison
    try:
        # Find grand total row
        grand_total_row = None
        for row in range(1, budget_ws.max_row + 1):
            v = budget_ws.cell(row=row, column=1).value
            if v and "grand" in str(v).lower() and "total" in str(v).lower():
                grand_total_row = row
                break

        comp4_score = 0.0

        if grand_total_row:
            # Check that grand total has a formula in column C
            gt_formula = budget_ws.cell(row=grand_total_row, column=3).value
            if gt_formula and isinstance(gt_formula, str) and ("SUM" in gt_formula.upper() or "+" in gt_formula):
                print(f"  Grand total formula found: {gt_formula}")
                comp4_score += 0.07
            elif gt_formula and isinstance(gt_formula, (int, float)):
                # Computed value present
                comp4_score += 0.05
                print(f"  Grand total value found: {gt_formula}")
            else:
                print(f"  FAIL: Grand total C column value: {gt_formula}")
        else:
            print(f"  FAIL: No 'GRAND TOTAL' row found")

        # Check budget tracking: look for $25,000 budget reference
        budget_found = False
        for row in range(1, min(budget_ws.max_row + 1, 10)):
            for col in range(1, budget_ws.max_column + 1):
                v = budget_ws.cell(row=row, column=col).value
                if v == 25000 or (isinstance(v, (int, float)) and abs(v - 25000) < 1):
                    budget_found = True
                    break
            if budget_found:
                break

        if budget_found:
            print(f"  Budget $25,000 reference found")
            comp4_score += 0.06
        else:
            print(f"  FAIL: Budget $25,000 not found")

        # Check for remaining/comparison formula (Budget vs Actual or remaining budget)
        budget_formula_found = False
        for row in range(1, min(budget_ws.max_row + 1, 10)):
            for col in range(1, budget_ws.max_column + 1):
                v = budget_ws.cell(row=row, column=col).value
                if v and isinstance(v, str) and "=" in v:
                    budget_formula_found = True
                    break
            if budget_formula_found:
                break

        # Also check if Budget vs Actual column exists (column I or similar)
        budget_vs_actual = False
        for row in range(budget_ws.max_row, max(budget_ws.max_row - 5, 1), -1):
            for col in range(1, budget_ws.max_column + 1):
                v = budget_ws.cell(row=row, column=col).value
                if v and isinstance(v, str) and "=" in v and ("/" in v or "B2" in v.upper() or "budget" in v.lower()):
                    budget_vs_actual = True
                    break
            if budget_vs_actual:
                break

        # Check vendor data rows for budget comparison formulas
        if not budget_vs_actual:
            for row in range(4, budget_ws.max_row + 1):
                last_col_val = budget_ws.cell(row=row, column=budget_ws.max_column).value
                if last_col_val and isinstance(last_col_val, str) and "=" in last_col_val:
                    budget_vs_actual = True
                    break

        if budget_vs_actual:
            print(f"  Budget vs Actual comparison formulas found")
            comp4_score += 0.07
        else:
            print(f"  FAIL: No budget vs actual comparison formulas found")

        if comp4_score > 0:
            print(f"PASS: Component 4 — Grand total + budget tracking ({comp4_score:.2f} pts)")
            total_score += comp4_score
        else:
            print(f"FAIL: Component 4 — No grand total or budget tracking found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Conditional formatting on Budget Tracker (0.10 points)
    # Task requires: yellow for <=30 days, red for <=7 days, using TODAY()
    try:
        cf_rules = list(budget_ws.conditional_formatting)
        has_cf = len(cf_rules) > 0

        has_today_formula = False
        has_two_rules = False
        rule_count = 0

        for cf in cf_rules:
            for rule in cf.rules:
                rule_count += 1
                if rule.formula:
                    formula_str = str(rule.formula).upper()
                    if "TODAY" in formula_str:
                        has_today_formula = True

        has_two_rules = rule_count >= 2

        comp5_score = 0.0
        if has_cf and has_today_formula:
            comp5_score += 0.05
        if has_two_rules and has_today_formula:
            comp5_score += 0.05

        if comp5_score > 0:
            print(f"PASS: Component 5 — Conditional formatting with {rule_count} rules using TODAY() ({comp5_score:.2f} pts)")
            total_score += comp5_score
        else:
            print(f"FAIL: Component 5 — Conditional formatting: has_cf={has_cf}, "
                  f"has_today={has_today_formula}, rules={rule_count}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Payment Timeline sheet with sorted vendor entries (0.15 points)
    try:
        if timeline_ws is None:
            print(f"FAIL: Component 6 — No Payment Timeline sheet found")
        else:
            comp6_score = 0.0

            # Check it has data rows (at least 5 vendor entries)
            data_rows = 0
            has_dates = False
            has_vendor_names = False

            # Find header row on timeline sheet
            tl_header_row = None
            for row in range(1, min(timeline_ws.max_row + 1, 8)):
                for col in range(1, timeline_ws.max_column + 1):
                    v = timeline_ws.cell(row=row, column=col).value
                    if v and "vendor" in str(v).lower():
                        tl_header_row = row
                        break
                if tl_header_row:
                    break

            start_row = (tl_header_row + 1) if tl_header_row else 2
            import datetime
            for row in range(start_row, timeline_ws.max_row + 1):
                # Check for date in first column
                date_val = timeline_ws.cell(row=row, column=1).value
                vendor_val = timeline_ws.cell(row=row, column=2).value
                if date_val is not None and vendor_val is not None:
                    data_rows += 1
                    if isinstance(date_val, datetime.datetime):
                        has_dates = True
                    if isinstance(vendor_val, str) and len(vendor_val) > 2:
                        has_vendor_names = True

            if data_rows >= 5:
                comp6_score += 0.05
                print(f"  Timeline has {data_rows} vendor entries")
            else:
                print(f"  FAIL: Timeline has only {data_rows} vendor entries (need >= 5)")

            if has_dates:
                comp6_score += 0.05
                print(f"  Timeline has date values")

            if has_vendor_names:
                comp6_score += 0.05
                print(f"  Timeline has vendor names")

            if comp6_score > 0:
                print(f"PASS: Component 6 — Payment Timeline with entries ({comp6_score:.2f} pts)")
                total_score += comp6_score
            else:
                print(f"FAIL: Component 6 — Insufficient timeline data")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: Conditional formatting on Payment Timeline (0.10 points)
    try:
        if timeline_ws is None:
            print(f"FAIL: Component 7 — No timeline sheet for CF check")
        else:
            tl_cf_rules = list(timeline_ws.conditional_formatting)
            tl_has_cf = len(tl_cf_rules) > 0
            tl_has_today = False
            tl_rule_count = 0

            for cf in tl_cf_rules:
                for rule in cf.rules:
                    tl_rule_count += 1
                    if rule.formula:
                        if "TODAY" in str(rule.formula).upper():
                            tl_has_today = True

            # Also accept if the timeline sheet has status formulas using TODAY()
            tl_has_status_formula = False
            for row in range(1, timeline_ws.max_row + 1):
                for col in range(1, timeline_ws.max_column + 1):
                    v = timeline_ws.cell(row=row, column=col).value
                    if v and isinstance(v, str) and "TODAY" in v.upper():
                        tl_has_status_formula = True
                        break
                if tl_has_status_formula:
                    break

            comp7_score = 0.0
            if tl_has_cf and tl_has_today:
                comp7_score = 0.10
                print(f"PASS: Component 7 — Timeline conditional formatting with {tl_rule_count} rules using TODAY() ({comp7_score:.2f} pts)")
            elif tl_has_status_formula:
                # Partial credit: has TODAY()-based status formulas even without CF
                comp7_score = 0.05
                print(f"PARTIAL: Component 7 — Timeline has TODAY() status formulas but no CF ({comp7_score:.2f} pts)")
            else:
                print(f"FAIL: Component 7 — No timeline conditional formatting or status formulas")

            total_score += comp7_score
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
