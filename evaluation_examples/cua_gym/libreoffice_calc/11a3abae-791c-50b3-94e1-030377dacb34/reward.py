"""
Reward Script: Financial summary formatting for income statement
Task ID: calc_gsd_014
Domain: libreoffice_calc
Scoring:
  Component 1: A1:B1 merged (0.15)
  Component 2: USD currency formatting on B5:B25 value cells (0.25)
  Component 3: Bold on header/subtotal rows (0.25)
  Component 4: Thick bottom border on B11 and B23 (0.20)
  Component 5: Light yellow background on row 25 (0.15)
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_014'


def persist_app_state(domain):
    """Try to save any unsaved LibreOffice state."""
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print(f"PERSIST: ctrl+s sent for {domain}")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: P&L sheet must exist
    if 'P&L' not in wb.sheetnames:
        print("CRITICAL: Sheet 'P&L' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['P&L']

    # Component 1: A1:B1 merged (0.15 points)
    try:
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        if any('A1:B1' in mr for mr in merged_ranges):
            print(f"PASS: Component 1 -- A1:B1 is merged (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 -- A1:B1 not merged. Merged ranges: {merged_ranges}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: USD currency formatting on B column value cells (0.25 points)
    # Check B6:B11, B14:B23, B25 -- cells that have numeric values
    try:
        currency_cells = []
        for row in list(range(6, 12)) + list(range(14, 24)) + [25]:
            cell = ws[f'B{row}']
            if not isinstance(cell, MergedCell) and cell.value is not None:
                currency_cells.append((f'B{row}', cell))

        if len(currency_cells) == 0:
            print(f"FAIL: Component 2 -- No value cells found in B column")
        else:
            formatted_count = 0
            for coord, cell in currency_cells:
                nf = cell.number_format
                # Accept formats containing $ and decimal places
                if nf and '$' in nf and '0.00' in nf:
                    formatted_count += 1
                else:
                    print(f"  INFO: {coord} number_format={nf!r} -- not USD currency")

            ratio = formatted_count / len(currency_cells)
            if ratio >= 0.9:
                print(f"PASS: Component 2 -- {formatted_count}/{len(currency_cells)} cells have USD currency format (0.25 pts)")
                total_score += 0.25
            elif ratio >= 0.5:
                partial = round(0.25 * ratio, 2)
                print(f"PARTIAL: Component 2 -- {formatted_count}/{len(currency_cells)} cells formatted ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 2 -- Only {formatted_count}/{len(currency_cells)} cells have USD currency format")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Bold on header/subtotal rows (0.25 points)
    # Rows 5(REVENUE), 11(Total Revenue), 13(EXPENSES), 23(Total Expenses), 25(NET INCOME)
    # Check both A and B columns for these rows (A must be bold; B must be bold for 11, 23, 25)
    try:
        bold_checks = {
            'A5': False, 'A11': False, 'A13': False, 'A23': False, 'A25': False,
            'B11': False, 'B23': False, 'B25': False,
        }
        for coord in bold_checks:
            cell = ws[coord]
            if not isinstance(cell, MergedCell) and cell.font.bold:
                bold_checks[coord] = True

        passed = sum(bold_checks.values())
        total_checks = len(bold_checks)
        ratio = passed / total_checks

        if ratio >= 0.9:
            print(f"PASS: Component 3 -- {passed}/{total_checks} bold checks passed (0.25 pts)")
            total_score += 0.25
        elif ratio >= 0.5:
            partial = round(0.25 * ratio, 2)
            print(f"PARTIAL: Component 3 -- {passed}/{total_checks} bold checks passed ({partial} pts)")
            total_score += partial
        else:
            failed = [k for k, v in bold_checks.items() if not v]
            print(f"FAIL: Component 3 -- Only {passed}/{total_checks} bold. Failed: {failed}")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Thick bottom border on B11 and B23 (0.20 points)
    try:
        borders_ok = 0
        for coord in ['B11', 'B23']:
            cell = ws[coord]
            if not isinstance(cell, MergedCell):
                bottom_style = cell.border.bottom.style
                if bottom_style in ('thick', 'medium'):
                    borders_ok += 1
                    print(f"  INFO: {coord} bottom border style={bottom_style}")
                else:
                    print(f"  INFO: {coord} bottom border style={bottom_style} -- expected thick")

        if borders_ok == 2:
            print(f"PASS: Component 4 -- Both B11 and B23 have thick bottom border (0.20 pts)")
            total_score += 0.20
        elif borders_ok == 1:
            print(f"PARTIAL: Component 4 -- 1/2 thick bottom borders (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 -- No thick bottom borders found on B11/B23")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # Component 5: Light yellow background (#FFFFCC) on row 25 (0.15 points)
    try:
        yellow_ok = 0
        for coord in ['A25', 'B25']:
            cell = ws[coord]
            if not isinstance(cell, MergedCell):
                try:
                    fill_rgb = cell.fill.fgColor.rgb
                except:
                    fill_rgb = None
                fill_type = cell.fill.fill_type

                # Accept FFFFFFCC (8-char ARGB with FF alpha) or 00FFFFCC
                if fill_rgb and 'FFFFCC' in str(fill_rgb) and fill_type == 'solid':
                    yellow_ok += 1
                    print(f"  INFO: {coord} fill={fill_rgb} type={fill_type}")
                else:
                    print(f"  INFO: {coord} fill={fill_rgb} type={fill_type} -- expected FFFFCC")

        if yellow_ok >= 1:
            print(f"PASS: Component 5 -- Row 25 has light yellow background (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 5 -- Row 25 does not have light yellow background")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
