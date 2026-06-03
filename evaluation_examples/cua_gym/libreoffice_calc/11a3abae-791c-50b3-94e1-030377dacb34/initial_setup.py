"""
Initial Setup: Create income statement spreadsheet with raw data
Task ID: calc_gsd_014
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_014'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "P&L"

    # Row 1: Report title (NOT merged in initial)
    ws["A1"] = "Income Statement"
    # Row 2: Period info
    ws["A2"] = "For the Year Ended December 31, 2025"
    # Row 3: blank
    # Row 4: blank (separator)

    # --- REVENUE SECTION (rows 5-11) ---
    ws["A5"] = "REVENUE"

    revenue_items = [
        ("Product Sales", 1245000),
        ("Service Revenue", 487500),
        ("Licensing Fees", 162000),
        ("Consulting Income", 98750),
        ("Subscription Revenue", 315000),
    ]
    for i, (label, amount) in enumerate(revenue_items, start=6):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=amount)

    ws["A11"] = "Total Revenue"
    ws["B11"] = 2308250  # sum of revenue items

    # Row 12: blank

    # --- EXPENSES SECTION (rows 13-23) ---
    ws["A13"] = "EXPENSES"

    expense_items = [
        ("Cost of Goods Sold", 498200),
        ("Salaries & Wages", 612000),
        ("Rent & Utilities", 84000),
        ("Marketing & Advertising", 137500),
        ("Software & Technology", 65400),
        ("Professional Services", 42800),
        ("Insurance", 28500),
        ("Depreciation", 53200),
        ("Travel & Entertainment", 31750),
    ]
    for i, (label, amount) in enumerate(expense_items, start=14):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=amount)

    ws["A23"] = "Total Expenses"
    ws["B23"] = 1553350  # sum of expense items

    # Row 24: blank

    # --- NET INCOME (row 25) ---
    ws["A25"] = "NET INCOME"
    ws["B25"] = 754900  # Total Revenue - Total Expenses

    # Set reasonable column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
