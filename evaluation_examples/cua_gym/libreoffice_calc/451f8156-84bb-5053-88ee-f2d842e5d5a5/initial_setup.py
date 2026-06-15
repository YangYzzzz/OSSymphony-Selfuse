"""
Initial Setup: Create annual sales review workbook with raw data
Task ID: calc_sales_095
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_095'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: AnnualData ---
    ws1 = wb.active
    ws1.title = 'AnnualData'

    # Headers
    headers = ['Rep', 'Deals Won', 'Deals Lost', 'Revenue', 'Avg Deal Size']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # Data rows
    data = [
        ['Alice', 24, 8, 680000],
        ['Bob', 18, 12, 520000],
        ['Carol', 30, 6, 850000],
        ['Dan', 15, 15, 390000],
        ['Eve', 22, 10, 610000],
        ['Frank', 28, 4, 780000],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # E2:E7 = D/B (Avg Deal Size formulas)
    for r in range(2, 8):
        ws1.cell(row=r, column=5, value=f'=D{r}/B{r}')

    # --- Sheet 2: Summary (metric labels only, NO formulas in B column) ---
    ws2 = wb.create_sheet('Summary')
    ws2['A1'] = 'Metric'
    ws2['B1'] = 'Value'

    metric_labels = [
        'Total Team Revenue',
        'Average Rep Revenue',
        'Median Rep Revenue',
        'Std Deviation',
        'Total Deals Won',
        'Total Deals Lost',
        'Team Win Rate',
        'Top Performer',
        'Bottom Performer',
    ]
    for r, label in enumerate(metric_labels, 2):
        ws2.cell(row=r, column=1, value=label)
    # B2:B10 intentionally left empty - that's the task

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
