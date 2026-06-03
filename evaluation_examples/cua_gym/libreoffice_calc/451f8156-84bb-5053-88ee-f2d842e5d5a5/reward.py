"""
Reward Script: Annual sales review workbook with summary statistics
Task ID: calc_sales_095
Domain: libreoffice_calc
Scoring: 7 components checking Summary sheet formulas (B2-B10) that were
         absent in initial_env and present in golden_env.
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_095'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def formula_contains(value, *keywords):
    """Check if a cell value is a formula string containing all given keywords (case-insensitive)."""
    if not isinstance(value, str) or not value.startswith('='):
        return False
    upper = value.upper()
    return all(kw.upper() in upper for kw in keywords)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    All scoring components verify formulas in the Summary sheet B column,
    which are empty in initial_env and populated in golden_env.
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Summary sheet must exist
    if 'Summary' not in wb.sheetnames:
        print("FAIL: 'Summary' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Summary']

    # Component 1: Total Team Revenue — B2 contains SUM formula referencing AnnualData D column (0.15 pts)
    try:
        val = ws['B2'].value
        if val is not None and formula_contains(str(val), 'SUM', 'D'):
            print(f"PASS: Component 1 — B2 has SUM formula for revenue: {val} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — Expected SUM formula for revenue in B2, found: {repr(val)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Average & Median Rep Revenue — B3 has AVERAGE, B4 has MEDIAN (0.15 pts)
    try:
        b3 = ws['B3'].value
        b4 = ws['B4'].value
        has_avg = b3 is not None and formula_contains(str(b3), 'AVERAGE')
        has_med = b4 is not None and formula_contains(str(b4), 'MEDIAN')
        if has_avg and has_med:
            print(f"PASS: Component 2 — B3 AVERAGE ({b3}), B4 MEDIAN ({b4}) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 — B3={repr(b3)} (avg:{has_avg}), B4={repr(b4)} (med:{has_med})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Std Deviation — B5 has STDEV formula (0.10 pts)
    try:
        val = ws['B5'].value
        if val is not None and formula_contains(str(val), 'STDEV'):
            print(f"PASS: Component 3 — B5 has STDEV formula: {val} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — Expected STDEV formula in B5, found: {repr(val)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Total Deals Won (B6) and Total Deals Lost (B7) — SUM formulas (0.15 pts)
    try:
        b6 = ws['B6'].value
        b7 = ws['B7'].value
        has_b6 = b6 is not None and formula_contains(str(b6), 'SUM', 'B')
        has_b7 = b7 is not None and formula_contains(str(b7), 'SUM', 'C')
        if has_b6 and has_b7:
            print(f"PASS: Component 4 — B6 SUM deals won ({b6}), B7 SUM deals lost ({b7}) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — B6={repr(b6)} (won:{has_b6}), B7={repr(b7)} (lost:{has_b7})")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Team Win Rate — B8 has a formula computing win rate (ratio involving B6 and B7) (0.15 pts)
    try:
        val = ws['B8'].value
        if val is not None and isinstance(val, str) and val.startswith('='):
            upper = val.upper()
            # Win rate formula should reference B6 and B7 (or equivalent SUM expressions)
            # and involve division
            if ('B6' in upper or 'SUM' in upper) and '/' in val:
                print(f"PASS: Component 5 — B8 has win rate formula: {val} (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 5 — B8 formula doesn't look like win rate: {val}")
        else:
            print(f"FAIL: Component 5 — Expected win rate formula in B8, found: {repr(val)}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Top Performer — B9 uses INDEX/MATCH with MAX to find top rep (0.15 pts)
    try:
        val = ws['B9'].value
        if val is not None and isinstance(val, str) and val.startswith('='):
            upper = val.upper()
            if 'MAX' in upper and ('INDEX' in upper or 'MATCH' in upper or 'VLOOKUP' in upper or 'XLOOKUP' in upper):
                print(f"PASS: Component 6 — B9 has top performer formula: {val} (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 6 — B9 formula missing INDEX/MATCH/MAX: {val}")
        else:
            print(f"FAIL: Component 6 — Expected INDEX/MATCH/MAX formula in B9, found: {repr(val)}")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: Bottom Performer — B10 uses INDEX/MATCH with MIN to find bottom rep (0.15 pts)
    try:
        val = ws['B10'].value
        if val is not None and isinstance(val, str) and val.startswith('='):
            upper = val.upper()
            if 'MIN' in upper and ('INDEX' in upper or 'MATCH' in upper or 'VLOOKUP' in upper or 'XLOOKUP' in upper):
                print(f"PASS: Component 7 — B10 has bottom performer formula: {val} (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 7 — B10 formula missing INDEX/MATCH/MIN: {val}")
        else:
            print(f"FAIL: Component 7 — Expected INDEX/MATCH/MIN formula in B10, found: {repr(val)}")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
