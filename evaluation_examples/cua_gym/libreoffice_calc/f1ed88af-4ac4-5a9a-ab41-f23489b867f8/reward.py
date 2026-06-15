"""
Reward Script: Delete sheets 'Draft_v1' and 'Draft_v2' but keep 'Draft_v3'. Then rename 'Draft_v3' to 'Final'.
Task ID: calc_ps_091
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): Draft_v1 and Draft_v2 are deleted
  Component 2 (0.2): Draft_v3 no longer exists (renamed away)
  Component 3 (0.2): 'Final' sheet exists
  Component 4 (0.3): 'Final' sheet contains original Draft_v3 data
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_091'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    sheet_names = wb.sheetnames
    print(f"INFO: Found sheets: {sheet_names}")

    # Component 1: Draft_v1 and Draft_v2 are deleted (0.3 points)
    # These sheets exist in initial_env but should be removed in golden_env
    try:
        draft_v1_gone = 'Draft_v1' not in sheet_names
        draft_v2_gone = 'Draft_v2' not in sheet_names
        if draft_v1_gone and draft_v2_gone:
            print(f"PASS: Component 1 — Both Draft_v1 and Draft_v2 deleted (0.3 pts)")
            total_score += 0.3
        elif draft_v1_gone or draft_v2_gone:
            # Partial: one of two deleted
            print(f"PARTIAL: Component 1 — Only one draft deleted. Draft_v1 gone={draft_v1_gone}, Draft_v2 gone={draft_v2_gone} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — Both Draft_v1 and Draft_v2 still exist")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Draft_v3 no longer exists as a sheet name (0.2 points)
    # It should have been renamed to 'Final', so 'Draft_v3' should not be present
    try:
        if 'Draft_v3' not in sheet_names:
            print(f"PASS: Component 2 — Draft_v3 no longer exists (renamed) (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 2 — Draft_v3 still exists in sheet list")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: 'Final' sheet exists (0.2 points)
    try:
        if 'Final' in sheet_names:
            print(f"PASS: Component 3 — 'Final' sheet exists (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — 'Final' sheet not found in {sheet_names}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: 'Final' sheet contains original Draft_v3 data (0.3 points)
    # Verify key cells from Draft_v3: headers and a few data cells
    try:
        if 'Final' in sheet_names:
            ws = wb['Final']
            checks_passed = 0
            total_checks = 5

            # Check header row
            if ws['A1'].value == 'Milestone':
                checks_passed += 1
            else:
                print(f"  DETAIL: A1 expected 'Milestone', found '{ws['A1'].value}'")

            if ws['B1'].value == 'Owner':
                checks_passed += 1
            else:
                print(f"  DETAIL: B1 expected 'Owner', found '{ws['B1'].value}'")

            if ws['F1'].value == 'Notes':
                checks_passed += 1
            else:
                print(f"  DETAIL: F1 expected 'Notes', found '{ws['F1'].value}'")

            # Check specific data cells from Draft_v3
            if ws['A2'].value == 'Requirements Gathering':
                checks_passed += 1
            else:
                print(f"  DETAIL: A2 expected 'Requirements Gathering', found '{ws['A2'].value}'")

            if ws['B3'].value == 'David Kim':
                checks_passed += 1
            else:
                print(f"  DETAIL: B3 expected 'David Kim', found '{ws['B3'].value}'")

            if checks_passed == total_checks:
                print(f"PASS: Component 4 — 'Final' sheet contains correct Draft_v3 data ({checks_passed}/{total_checks} checks) (0.3 pts)")
                total_score += 0.3
            elif checks_passed > 0:
                partial = round(0.3 * checks_passed / total_checks, 2)
                print(f"PARTIAL: Component 4 — {checks_passed}/{total_checks} data checks passed ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 4 — No data checks passed for 'Final' sheet")
        else:
            print(f"FAIL: Component 4 — Cannot check data, 'Final' sheet does not exist")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
