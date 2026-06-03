"""
Initial Setup: Delete sheets 'Draft_v1' and 'Draft_v2' but keep 'Draft_v3'. Then rename 'Draft_v3' to 'Final'.
Task ID: calc_ps_091
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_091'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Common styles ---
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    def style_header(ws, headers, row=1):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    # --- Sheet 1: Main ---
    ws_main = wb.active
    ws_main.title = 'Main'

    main_headers = ['Project', 'Status', 'Lead', 'Budget ($)', 'Deadline']
    style_header(ws_main, main_headers)

    main_data = [
        ['Website Redesign', 'In Progress', 'Sarah Chen', 125000, '2025-06-30'],
        ['Mobile App v2', 'Planning', 'Marcus Johnson', 89000, '2025-09-15'],
        ['Data Pipeline', 'Completed', 'Aisha Patel', 67500, '2025-03-01'],
        ['CRM Integration', 'In Progress', 'Carlos Rivera', 45000, '2025-07-20'],
        ['Security Audit', 'Not Started', 'Emily Watson', 32000, '2025-08-10'],
        ['Cloud Migration', 'In Progress', 'David Kim', 210000, '2025-12-31'],
        ['API Gateway', 'Planning', 'Fatima Al-Hassan', 55000, '2025-10-15'],
        ['Analytics Dashboard', 'In Progress', 'James O\'Brien', 78000, '2025-05-30'],
        ['DevOps Tooling', 'Completed', 'Priya Sharma', 41000, '2025-02-28'],
        ['Customer Portal', 'Planning', 'Lucas Andersen', 96000, '2025-11-30'],
        ['ML Pipeline', 'Not Started', 'Nina Kowalski', 150000, '2026-01-31'],
        ['Compliance System', 'In Progress', 'Omar Hassan', 63000, '2025-08-31'],
    ]
    for r, row_data in enumerate(main_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_main.cell(row=r, column=c, value=val)
            cell.border = thin_border

    ws_main.column_dimensions['A'].width = 22
    ws_main.column_dimensions['B'].width = 14
    ws_main.column_dimensions['C'].width = 18
    ws_main.column_dimensions['D'].width = 14
    ws_main.column_dimensions['E'].width = 14

    # --- Sheet 2: Draft_v1 ---
    ws_d1 = wb.create_sheet('Draft_v1')
    d1_headers = ['Employee', 'Department', 'Q1 Sales', 'Q2 Sales', 'Region']
    style_header(ws_d1, d1_headers)

    d1_data = [
        ['Rachel Green', 'Sales', 45230, 51200, 'Northeast'],
        ['Tom Bradley', 'Sales', 38900, 42100, 'Southeast'],
        ['Linda Park', 'Sales', 52100, 48700, 'West'],
        ['Steve Morris', 'Sales', 41500, 45600, 'Midwest'],
        ['Angela Torres', 'Sales', 47800, 50300, 'Southwest'],
        ['Paul Henderson', 'Sales', 36200, 39800, 'Northeast'],
        ['Diana Chen', 'Sales', 55400, 58900, 'West'],
        ['Robert Taylor', 'Sales', 43700, 46200, 'Southeast'],
        ['Maria Santos', 'Sales', 49100, 52800, 'Midwest'],
        ['Kevin O\'Malley', 'Sales', 37600, 41300, 'Northeast'],
        ['Jennifer Wu', 'Sales', 51800, 54700, 'West'],
        ['Brian Cooper', 'Sales', 40200, 43900, 'Southwest'],
    ]
    for r, row_data in enumerate(d1_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_d1.cell(row=r, column=c, value=val)
            cell.border = thin_border

    ws_d1.column_dimensions['A'].width = 18
    ws_d1.column_dimensions['B'].width = 14
    ws_d1.column_dimensions['C'].width = 12
    ws_d1.column_dimensions['D'].width = 12
    ws_d1.column_dimensions['E'].width = 14

    # --- Sheet 3: Draft_v2 ---
    ws_d2 = wb.create_sheet('Draft_v2')
    d2_headers = ['Product', 'Category', 'Units Sold', 'Revenue ($)', 'Margin (%)']
    style_header(ws_d2, d2_headers)

    d2_data = [
        ['ProWidget X100', 'Hardware', 1250, 187500, 32.5],
        ['SmartSense Hub', 'IoT', 890, 133500, 41.2],
        ['DataFlow Pro', 'Software', 2100, 315000, 68.7],
        ['CloudSync Elite', 'Software', 1780, 267000, 72.1],
        ['SecureVault Plus', 'Security', 950, 142500, 55.3],
        ['NetBoost 5G', 'Hardware', 620, 93000, 28.9],
        ['AI Analyzer', 'Software', 1400, 210000, 65.4],
        ['EdgeCompute Mini', 'Hardware', 830, 124500, 35.8],
        ['TeamConnect Pro', 'Software', 2450, 367500, 71.6],
        ['BioScan ID', 'Security', 540, 81000, 48.2],
        ['StreamLine HD', 'Hardware', 1100, 165000, 30.7],
        ['AutoPilot Suite', 'Software', 1670, 250500, 69.3],
    ]
    for r, row_data in enumerate(d2_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_d2.cell(row=r, column=c, value=val)
            cell.border = thin_border

    ws_d2.column_dimensions['A'].width = 20
    ws_d2.column_dimensions['B'].width = 14
    ws_d2.column_dimensions['C'].width = 12
    ws_d2.column_dimensions['D'].width = 14
    ws_d2.column_dimensions['E'].width = 12

    # --- Sheet 4: Draft_v3 ---
    ws_d3 = wb.create_sheet('Draft_v3')
    d3_headers = ['Milestone', 'Owner', 'Start Date', 'End Date', 'Completion (%)', 'Notes']
    style_header(ws_d3, d3_headers)

    d3_data = [
        ['Requirements Gathering', 'Sarah Chen', '2025-01-06', '2025-01-31', 100, 'Stakeholder interviews done'],
        ['Architecture Design', 'David Kim', '2025-02-03', '2025-02-28', 100, 'Approved by CTO'],
        ['Backend Development', 'Marcus Johnson', '2025-03-03', '2025-04-25', 85, 'API endpoints in review'],
        ['Frontend Development', 'Priya Sharma', '2025-03-10', '2025-05-02', 70, 'Dashboard component pending'],
        ['Database Migration', 'Aisha Patel', '2025-04-01', '2025-04-30', 60, 'Schema finalized'],
        ['Integration Testing', 'Carlos Rivera', '2025-05-01', '2025-05-30', 25, 'Test plan drafted'],
        ['Performance Testing', 'Emily Watson', '2025-05-15', '2025-06-13', 10, 'Baseline metrics captured'],
        ['Security Review', 'Omar Hassan', '2025-06-02', '2025-06-20', 0, 'Pending pen test schedule'],
        ['User Acceptance Testing', 'Fatima Al-Hassan', '2025-06-16', '2025-07-04', 0, 'UAT scripts ready'],
        ['Deployment Planning', 'James O\'Brien', '2025-06-23', '2025-07-11', 0, 'Runbook in draft'],
        ['Production Rollout', 'David Kim', '2025-07-14', '2025-07-18', 0, 'Blue-green deploy strategy'],
        ['Post-Launch Monitoring', 'Lucas Andersen', '2025-07-21', '2025-08-01', 0, 'Alerts configured'],
    ]
    for r, row_data in enumerate(d3_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_d3.cell(row=r, column=c, value=val)
            cell.border = thin_border

    ws_d3.column_dimensions['A'].width = 24
    ws_d3.column_dimensions['B'].width = 18
    ws_d3.column_dimensions['C'].width = 14
    ws_d3.column_dimensions['D'].width = 14
    ws_d3.column_dimensions['E'].width = 16
    ws_d3.column_dimensions['F'].width = 30

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
