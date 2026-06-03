"""
Reward Script: Fix shifted relative references in Sheet2 C2:C50
Task ID: calc_tbl_035
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): No formulas in Sheet2 C2:C50 contain cross-sheet "Sheet1!" references
  Component 2 (0.4): All 49 formulas exactly match =A{row}*B{row} pattern
  Component 3 (0.2): Sheet1 formulas unchanged (data integrity gate on task change)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_035'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sheet2 must exist
    if 'Sheet2' not in wb.sheetnames:
        print("FAIL: Sheet2 not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws2 = wb['Sheet2']

    # Component 1: No cross-sheet references in Sheet2 C2:C50 (0.4 points)
    # In the initial file, ALL 49 formulas have "Sheet1!" references.
    # The task requires removing these cross-sheet refs.
    try:
        cross_sheet_count = 0
        for r in range(2, 51):
            val = ws2.cell(row=r, column=3).value
            if val is not None and 'sheet1!' in str(val).lower():
                cross_sheet_count += 1

        if cross_sheet_count == 0:
            print(f"PASS: Component 1 — No cross-sheet references found in Sheet2 C2:C50 (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — {cross_sheet_count}/49 formulas still have Sheet1! references")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All formulas match =A{row}*B{row} exactly (0.4 points)
    # Award partial credit proportional to how many formulas are correct.
    try:
        correct_count = 0
        total_formulas = 49  # C2:C50

        for r in range(2, 51):
            val = ws2.cell(row=r, column=3).value
            if val is None:
                continue
            expected = f'=A{r}*B{r}'
            # Normalize: strip spaces and compare case-insensitively
            actual_norm = str(val).upper().replace(' ', '')
            expected_norm = expected.upper().replace(' ', '')
            if actual_norm == expected_norm:
                correct_count += 1

        if correct_count == total_formulas:
            print(f"PASS: Component 2 — All {total_formulas} formulas match =A{{row}}*B{{row}} (0.4 pts)")
            total_score += 0.4
        elif correct_count > 0:
            partial = round(0.4 * (correct_count / total_formulas), 4)
            print(f"PARTIAL: Component 2 — {correct_count}/{total_formulas} formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — 0/{total_formulas} formulas match expected pattern")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Sheet1 formulas unchanged — data integrity (0.2 points)
    # Sheet1 C2:C50 should still have =A{row}*B{row} (the original correct formulas).
    # This checks the task didn't accidentally break Sheet1.
    # In initial_env, Sheet1 already has these formulas, so this alone would pass on initial.
    # BUT we gate it: only award if Component 1 also passed (cross-sheet refs removed).
    # This ensures Component 3 only scores when the task change is present.
    try:
        if 'Sheet1' not in wb.sheetnames:
            print("FAIL: Component 3 — Sheet1 not found")
        else:
            ws1 = wb['Sheet1']
            sheet1_ok = 0
            for r in range(2, 51):
                val = ws1.cell(row=r, column=3).value
                expected = f'=A{r}*B{r}'
                if val is not None:
                    actual_norm = str(val).upper().replace(' ', '')
                    expected_norm = expected.upper().replace(' ', '')
                    if actual_norm == expected_norm:
                        sheet1_ok += 1

            # Only award points if cross-sheet refs are already removed (Component 1 passed)
            if cross_sheet_count == 0 and sheet1_ok == 49:
                print(f"PASS: Component 3 — Sheet1 formulas intact, {sheet1_ok}/49 correct (0.2 pts)")
                total_score += 0.2
            elif cross_sheet_count > 0:
                print(f"FAIL: Component 3 — Gated on Component 1 (cross-sheet refs still present)")
            else:
                print(f"FAIL: Component 3 — Sheet1 has {sheet1_ok}/49 correct formulas (expected 49)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
