"""
Initial Setup: Fix incorrectly shifted formulas in Sheet2 after cross-sheet paste
Task ID: calc_tbl_035
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_035'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: Correct product pricing calculations ---
    ws1 = wb.active
    ws1.title = 'Sheet1'

    # Headers
    headers = ['Quantity', 'Unit Price', 'Total Cost']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # Realistic product order data for 49 rows (rows 2-50)
    import random
    random.seed(42)

    quantities = [
        15, 8, 22, 5, 30, 12, 18, 3, 25, 10,
        7, 45, 20, 14, 9, 35, 6, 28, 11, 16,
        40, 2, 33, 19, 27, 13, 50, 4, 21, 17,
        38, 23, 1, 31, 26, 8, 42, 15, 36, 24,
        29, 37, 10, 34, 48, 6, 19, 14, 32,
    ]

    unit_prices = [
        12.50, 45.00, 8.75, 120.00, 3.25, 67.50, 15.00, 250.00, 5.50, 34.00,
        89.99, 2.75, 18.50, 42.00, 95.00, 7.25, 155.00, 4.00, 62.50, 23.75,
        1.50, 310.00, 9.00, 38.50, 6.75, 55.00, 2.25, 185.00, 11.00, 47.50,
        3.50, 14.25, 475.00, 5.00, 8.25, 72.00, 2.00, 29.50, 4.75, 16.00,
        10.50, 3.75, 58.00, 6.25, 1.75, 112.00, 22.50, 43.00, 7.50,
    ]

    for i in range(49):
        r = i + 2
        ws1.cell(row=r, column=1, value=quantities[i])
        ws1.cell(row=r, column=2, value=unit_prices[i])
        # Correct formula: =A<row>*B<row>
        ws1.cell(row=r, column=3, value=f'=A{r}*B{r}')

    # --- Sheet2: Same data but with BROKEN formulas in column C ---
    # Simulates incorrect paste where references shifted (pointing to Sheet1 instead of Sheet2)
    ws2 = wb.create_sheet('Sheet2')

    # Same headers
    for col, h in enumerate(headers, 1):
        ws2.cell(row=1, column=col, value=h)

    # Same data in columns A and B
    for i in range(49):
        r = i + 2
        ws2.cell(row=r, column=1, value=quantities[i])
        ws2.cell(row=r, column=2, value=unit_prices[i])

        # BROKEN formulas: reference Sheet1 instead of local Sheet2 cells
        # This is what happens when you paste from Sheet1 to Sheet2 incorrectly
        ws2.cell(row=r, column=3, value=f'=Sheet1!A{r}*Sheet1!B{r}')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
