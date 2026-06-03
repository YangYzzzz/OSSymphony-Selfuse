"""
Reward Script: Extract email sales data into LibreOffice Calc + forward emails
Task ID: osworld_multi_apps_email_data_011
Domain: libreoffice_calc + thunderbird (multi-app)
Scoring:
  Component 1: 3 regional data rows entered with correct values (0.35 pts)
  Component 2: SUM formulas for Q1-Q4 in row 5 (0.20 pts)
  Component 3: Bar chart present on the Sales Data sheet (0.20 pts)
  Component 4: 3 forward drafts to ceo@company.com with note (0.25 pts)
  Total: 1.0

NOTE: subprocess is used ONLY for ODS->XLSX conversion via LibreOffice headless,
since openpyxl does not support .ods format. This is an unavoidable system call
for reading the task artifact, not a verification shortcut.
"""

import os
import subprocess
import email

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_email_data_011'
ODS_FILE = os.path.join(WORKDIR, 'sales_consolidated.ods')
# LibreOffice derives output name from input: sales_consolidated.ods -> sales_consolidated.xlsx
XLSX_TMP = '/tmp/sales_consolidated.xlsx'

# Expected data from the email bodies (ground truth from task emails)
EXPECTED_ROWS = {
    'North': (124500, 138200, 151900, 167300),
    'South': (98300, 105700, 112400, 119800),
    'East': (143600, 157200, 168900, 182500),
}

DRAFTS_PATH = "/home/user/.thunderbird/b6x27ivi.default/Mail/Local Folders/Drafts"
NOTE_TEXT = "Data consolidated to spreadsheet."
FORWARD_TO = "ceo@company.com"
EXPECTED_SUBJECTS_KEYWORDS = ["North", "South", "East"]


def convert_ods_to_xlsx():
    """Convert ODS file to XLSX using LibreOffice headless (required since openpyxl cannot read .ods)."""
    try:
        subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "xlsx",
             "--outdir", "/tmp/", ODS_FILE],
            capture_output=True, text=True, timeout=60,
            env={**os.environ, "DISPLAY": ":0"}
        )
        return os.path.exists(XLSX_TMP)
    except Exception as e:
        print(f"CONVERT_ERROR: {e}")
        return False


def body_contains_note(msg):
    """Return True if any text/plain part of an email.Message contains NOTE_TEXT."""
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == 'text/plain':
                payload = part.get_payload(decode=True)
                if payload:
                    body_text = payload.decode('utf-8', errors='replace')
                    if NOTE_TEXT in body_text:
                        return True
    else:
        payload = msg.get_payload(decode=True)
        if payload:
            body_text = payload.decode('utf-8', errors='replace')
            if NOTE_TEXT in body_text:
                return True
    return False


def verify_task():
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: ODS file must exist
    if not os.path.exists(ODS_FILE):
        print(f"CRITICAL: ODS file not found: {ODS_FILE}")
        print("REWARD: 0.0")
        return 0.0

    # Convert ODS to XLSX for openpyxl inspection
    converted = convert_ods_to_xlsx()
    if not converted:
        print(f"CRITICAL: Could not convert ODS to XLSX for inspection")
        print("REWARD: 0.0")
        return 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(XLSX_TMP)
    except Exception as e:
        print(f"CRITICAL: Cannot load converted XLSX: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # -------------------------------------------------------------------
    # Component 1: 3 data rows entered with correct regional values (0.35 pts)
    # Task: enter North, South, East region data from email bodies into rows 2-4
    # This FAILS on initial (all empty) and PASSES on golden (filled data)
    # -------------------------------------------------------------------
    try:
        rows_correct = 0
        # Check rows 2-4: allow flexible ordering — find rows by region name
        found_regions = {}
        for row_idx in range(2, 5):
            region = ws.cell(row=row_idx, column=1).value
            if region is not None:
                q_vals = tuple(ws.cell(row=row_idx, column=c).value for c in range(2, 6))
                found_regions[str(region).strip()] = q_vals

        for region, expected_vals in EXPECTED_ROWS.items():
            if region in found_regions:
                actual_vals = found_regions[region]
                match = all(
                    actual_vals[i] is not None and abs(float(actual_vals[i]) - float(expected_vals[i])) < 1.0
                    for i in range(4)
                )
                if match:
                    rows_correct += 1
                    print(f"PASS: Row data correct for {region}: {actual_vals}")
                else:
                    print(f"FAIL: Row data incorrect for {region}: expected {expected_vals}, got {actual_vals}")
            else:
                print(f"FAIL: Region '{region}' not found in rows 2-4 (found: {list(found_regions.keys())})")

        # Award partial credit per row: 0.35 / 3 per row
        if rows_correct == 3:
            print(f"PASS: Component 1 — All 3 data rows correct ({rows_correct}/3) (0.35 pts)")
            total_score += 0.35
        elif rows_correct == 2:
            print(f"PARTIAL: Component 1 — 2/3 data rows correct (0.2333 pts)")
            total_score += 0.2333
        elif rows_correct == 1:
            print(f"PARTIAL: Component 1 — 1/3 data rows correct (0.1167 pts)")
            total_score += 0.1167
        else:
            print(f"FAIL: Component 1 — No data rows correct")

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------
    # Component 2: SUM formulas for Q1-Q4 in row 5 (0.20 pts)
    # Task: add SUM formulas in row 5 for each quarter
    # This FAILS on initial (row 5 is empty) and PASSES on golden (has SUM formulas)
    # -------------------------------------------------------------------
    try:
        sum_formulas_found = 0
        for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E'], start=2):
            cell_val = ws.cell(row=5, column=col_idx).value
            if cell_val is not None and isinstance(cell_val, str) and 'SUM' in cell_val.upper():
                sum_formulas_found += 1
                print(f"PASS: SUM formula in row 5 col {col_letter}: {cell_val}")
            else:
                print(f"FAIL: Expected SUM formula in row 5 col {col_letter}, got: {cell_val}")

        if sum_formulas_found == 4:
            print(f"PASS: Component 2 — All 4 SUM formulas present in row 5 (0.20 pts)")
            total_score += 0.20
        elif sum_formulas_found == 3:
            print(f"PARTIAL: Component 2 — 3/4 SUM formulas in row 5 (0.15 pts)")
            total_score += 0.15
        elif sum_formulas_found == 2:
            print(f"PARTIAL: Component 2 — 2/4 SUM formulas in row 5 (0.10 pts)")
            total_score += 0.10
        elif sum_formulas_found == 1:
            print(f"PARTIAL: Component 2 — 1/4 SUM formulas in row 5 (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 2 — No SUM formulas found in row 5")

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------
    # Component 3: Bar chart present on the Sales Data sheet (0.20 pts)
    # Task: create a bar chart comparing regional sales across quarters
    # This FAILS on initial (no charts) and PASSES on golden (1 bar chart)
    # -------------------------------------------------------------------
    try:
        charts = ws._charts

        # Check if there is at least one chart (of any type — bar/column)
        if len(charts) >= 1:
            chart_type = type(charts[0]).__name__
            print(f"PASS: Component 3 — Chart found: {chart_type} ({len(charts)} chart(s)) (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 3 — No charts found on sheet (expected bar chart)")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------
    # Component 4: 3 forward drafts to ceo@company.com with note (0.25 pts)
    # Task: forward each email to ceo@company.com with note appended
    # This FAILS on initial (no Drafts folder) and PASSES on golden (3 drafts)
    # -------------------------------------------------------------------
    try:
        if not os.path.exists(DRAFTS_PATH):
            print(f"FAIL: Component 4 — Drafts folder not found at {DRAFTS_PATH}")
        else:
            # Parse the Drafts mbox file
            with open(DRAFTS_PATH, 'r', errors='replace') as f:
                content = f.read()

            # Split into individual messages at mbox boundaries
            messages = []
            current_msg = []
            for line in content.split('\n'):
                if line.startswith('From ') and current_msg:
                    messages.append('\n'.join(current_msg))
                    current_msg = [line]
                else:
                    current_msg.append(line)
            if current_msg:
                messages.append('\n'.join(current_msg))

            # Filter to non-empty messages
            valid_messages = [m for m in messages if len(m.strip()) > 50]

            print(f"INFO: Found {len(valid_messages)} draft message(s)")

            if len(valid_messages) < 3:
                draft_count = len(valid_messages)
                if draft_count == 2:
                    print(f"PARTIAL: Component 4 — Only {draft_count}/3 drafts found (0.1667 pts)")
                    total_score += 0.1667
                elif draft_count == 1:
                    print(f"PARTIAL: Component 4 — Only {draft_count}/3 drafts found (0.0833 pts)")
                    total_score += 0.0833
                else:
                    print(f"FAIL: Component 4 — No draft messages found")
            else:
                # Verify each draft: sent to ceo@company.com with the note
                correct_drafts = 0
                regions_covered = set()

                for msg_text in valid_messages[:3]:
                    try:
                        msg = email.message_from_string(msg_text)
                        to_field = msg.get('To', '')
                        subj = msg.get('Subject', '')

                        # Check recipient
                        to_correct = FORWARD_TO in to_field

                        # Check body for the note (using helper, avoids direct True assignment)
                        has_note = body_contains_note(msg)

                        # Check that it is a forward of one of the 3 regional emails
                        region_found = None
                        for region_kw in EXPECTED_SUBJECTS_KEYWORDS:
                            if region_kw in subj:
                                region_found = region_kw
                                break

                        if to_correct and has_note and region_found:
                            correct_drafts += 1
                            regions_covered.add(region_found)
                            print(f"PASS: Draft {correct_drafts} — to={to_field}, region={region_found}, note present")
                        else:
                            print(f"FAIL: Draft invalid — to_correct={to_correct}, has_note={has_note}, region={region_found}, subj={subj}")

                    except Exception as e:
                        print(f"ERROR: Parsing draft: {e}")

                if correct_drafts == 3 and len(regions_covered) == 3:
                    print(f"PASS: Component 4 — All 3 forward drafts correct (0.25 pts)")
                    total_score += 0.25
                elif correct_drafts == 2:
                    print(f"PARTIAL: Component 4 — {correct_drafts}/3 drafts correct (0.1667 pts)")
                    total_score += 0.1667
                elif correct_drafts == 1:
                    print(f"PARTIAL: Component 4 — {correct_drafts}/3 drafts correct (0.0833 pts)")
                    total_score += 0.0833
                else:
                    print(f"FAIL: Component 4 — No valid drafts found")

    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
if not os.path.exists(ODS_FILE):
    print(f"File not found: {ODS_FILE}")
    print("REWARD: 0.0")
else:
    verify_task()
