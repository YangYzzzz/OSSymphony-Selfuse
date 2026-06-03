"""
Initial Setup: Product inventory table with monthly units sold — no total row, no chart
Task ID: osworld_calc_multi_chart_computed_003
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_multi_chart_computed_003'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Inventory ---
    ws = wb.active
    ws.title = 'Inventory'

    # Column headers: Product | Jan | Feb | Mar | Apr | May | Jun
    headers = ['Product', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Realistic product inventory data (10 products, units sold Jan-Jun)
    data = [
        ['Wireless Headphones',   1240, 1380, 1190, 1450, 1620, 1780],
        ['USB-C Hub (7-port)',     890,  920,  870,  1010, 1130, 1095],
        ['Mechanical Keyboard',   530,  610,  580,  640,  720,  695],
        ['27" Monitor 4K',        310,  285,  330,  370,  410,  395],
        ['Laptop Stand Aluminum', 760,  830,  795,  870,  920,  985],
        ['Webcam 1080p',          415,  480,  510,  550,  490,  525],
        ['Ergonomic Mouse',       680,  710,  665,  730,  780,  815],
        ['HDMI Cable 2m',        1050, 1120,  980, 1090, 1150, 1210],
        ['Desk LED Lamp',         390,  420,  445,  410,  475,  500],
        ['Power Strip 6-outlet',  570,  605,  590,  625,  670,  715],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 1:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='right')

    # NOTE: No total row — that is what the agent must add.
    # NOTE: No chart — that is what the agent must create.

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 28
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup — open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
