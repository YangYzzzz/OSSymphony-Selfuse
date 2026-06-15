"""
Reward Script: Add total row with SUM formulas and create a line chart titled 'Units Sold Trend'
Task ID: osworld_calc_multi_chart_computed_003
Domain: libreoffice_calc
Scoring:
  - Component 1: Total row label 'Total' in A12        (0.2 pts)
  - Component 2: SUM formulas B12:G12 referencing B2:B11 through G2:G11  (0.4 pts)
  - Component 3: A line chart exists in the sheet       (0.2 pts)
  - Component 4: Chart title is 'Units Sold Trend'      (0.2 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_multi_chart_computed_003'


def extract_chart_title(chart):
    """
    Attempt to extract the chart title string from an openpyxl chart object.
    Returns the title string or None if it cannot be found.
    """
    try:
        if chart.title is None:
            return None
        # Try rich text extraction: title.tx.rich.p[0].r[0].t
        title_obj = chart.title
        paragraphs = title_obj.tx.rich.p
        if paragraphs:
            runs = paragraphs[0].r
            if runs:
                return runs[0].t
    except Exception:
        pass

    try:
        # Try strRef approach
        title_obj = chart.title
        strref = title_obj.tx.strRef
        if strref:
            return str(strref)
    except Exception:
        pass

    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Inventory' sheet must exist
    if 'Inventory' not in wb.sheetnames:
        print("CRITICAL: 'Inventory' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Inventory']

    # Component 1: Total row label 'Total' in A12 (0.2 points)
    # This must FAIL on initial_env (no row 12) and PASS on golden_env
    try:
        a12_val = ws.cell(row=12, column=1).value
        if a12_val is not None and str(a12_val).strip().lower() == 'total':
            print(f"PASS: Component 1 — A12 contains 'Total' label (value: {repr(a12_val)}) (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 1 — A12 expected 'Total', found: {repr(a12_val)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: SUM formulas in B12:G12 referencing B2:B11 through G2:G11 (0.4 points)
    # Each month column should have a SUM formula covering the 10 product rows.
    # Award 0.4 only if all 6 formulas are present and correct.
    # This must FAIL on initial_env and PASS on golden_env.
    try:
        columns = ['B', 'C', 'D', 'E', 'F', 'G']
        col_numbers = [2, 3, 4, 5, 6, 7]
        # Expected pattern: =SUM(X2:X11) for each column letter X
        sum_formula_count = 0
        formula_details = []
        for col_letter, col_num in zip(columns, col_numbers):
            cell_val = ws.cell(row=12, column=col_num).value
            if cell_val is None:
                formula_details.append(f"{col_letter}12: None")
                continue
            val_str = str(cell_val).strip().upper().replace(' ', '')
            expected = f"=SUM({col_letter}2:{col_letter}11)"
            if val_str == expected.upper():
                sum_formula_count += 1
                formula_details.append(f"{col_letter}12: OK ({cell_val})")
            else:
                formula_details.append(f"{col_letter}12: got {repr(cell_val)}, expected {expected}")

        if sum_formula_count == 6:
            print(f"PASS: Component 2 — All 6 SUM formulas present in B12:G12 (0.4 pts)")
            print(f"  Details: {', '.join(formula_details)}")
            total_score += 0.4
        elif sum_formula_count > 0:
            # Partial credit within this component: proportional, but we round to 2 decimal places
            partial = round((sum_formula_count / 6) * 0.4, 2)
            print(f"PARTIAL: Component 2 — {sum_formula_count}/6 SUM formulas correct ({partial} pts)")
            print(f"  Details: {', '.join(formula_details)}")
            if partial > 0:
                total_score += partial
        else:
            print(f"FAIL: Component 2 — No SUM formulas found in B12:G12")
            print(f"  Details: {', '.join(formula_details)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: A line chart exists in the sheet (0.2 points)
    # This must FAIL on initial_env (no charts) and PASS on golden_env (1 chart).
    try:
        charts = ws._charts
        line_charts = [c for c in charts if type(c).__name__ == 'LineChart']
        if len(line_charts) >= 1:
            print(f"PASS: Component 3 — Line chart found ({len(line_charts)} line chart(s)) (0.2 pts)")
            total_score += 0.2
        elif len(charts) >= 1:
            # Some chart exists but not a line chart — partial credit
            print(f"PARTIAL: Component 3 — Chart exists but not a LineChart (type: {type(charts[0]).__name__}). Awarding 0.1 pts")
            total_score += 0.1
        else:
            print(f"FAIL: Component 3 — No charts found in Inventory sheet (0 pts)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Chart title is 'Units Sold Trend' (0.2 points)
    # This must FAIL on initial_env (no chart) and PASS on golden_env.
    try:
        charts = ws._charts
        if len(charts) >= 1:
            chart = charts[0]
            title_text = extract_chart_title(chart)
            if title_text is not None and title_text.strip() == 'Units Sold Trend':
                print(f"PASS: Component 4 — Chart title is 'Units Sold Trend' (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 4 — Expected chart title 'Units Sold Trend', found: {repr(title_text)}")
        else:
            print(f"FAIL: Component 4 — No chart found; cannot check title")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
