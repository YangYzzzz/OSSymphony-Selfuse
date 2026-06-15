"""
Reward Script: Invoice Aging Report — Add Days Outstanding column with conditional formatting and frozen header
Task ID: calc_fin_invoice_aging_003
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): F1 header = 'Days Outstanding'
  Component 2 (0.35): F2:F45 all contain correct IF(status=Paid,0,TODAY()-date) formula
  Component 3 (0.25): Conditional formatting on F2:F45 with >90 -> red (#FF0000) and >60 -> orange (#FFA500)
  Component 4 (0.15): Row 1 frozen (freeze_panes == 'A2')
Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_invoice_aging_003'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Invoices' sheet must exist
    if 'Invoices' not in wb.sheetnames:
        print("CRITICAL: 'Invoices' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Invoices']

    # Component 1: F1 header is 'Days Outstanding' (0.25 points)
    try:
        f1_value = ws.cell(row=1, column=6).value
        if f1_value is not None and str(f1_value).strip() == 'Days Outstanding':
            print(f"PASS: Component 1 — F1 header = 'Days Outstanding' (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — expected F1='Days Outstanding', found: {repr(f1_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: F2:F45 all contain correct IF formula (0.35 points)
    # Expected pattern: =IF(E{row}="Paid",0,TODAY()-C{row})  (case-insensitive, spaces optional)
    try:
        formula_pattern = re.compile(
            r'^=IF\(E(\d+)="Paid",\s*0,\s*TODAY\(\)-C(\d+)\)$',
            re.IGNORECASE
        )
        correct_count = 0
        total_data_rows = 44  # rows 2-45
        wrong_formulas = []

        for row in range(2, 46):
            cell_val = ws.cell(row=row, column=6).value
            if cell_val is None:
                wrong_formulas.append(f"F{row}: None")
                continue
            val_str = str(cell_val).strip().replace(' ', '')
            # Check pattern with correct row references
            m = formula_pattern.match(val_str)
            if m and m.group(1) == str(row) and m.group(2) == str(row):
                correct_count += 1
            else:
                wrong_formulas.append(f"F{row}: {repr(cell_val)}")

        if correct_count == total_data_rows:
            print(f"PASS: Component 2 — All {total_data_rows} IF formulas in F2:F45 are correct (0.35 pts)")
            total_score += 0.35
        elif correct_count >= total_data_rows * 0.9:
            # Partial credit: at least 90% correct
            partial = 0.20
            print(f"PARTIAL: Component 2 — {correct_count}/{total_data_rows} IF formulas correct ({partial} pts)")
            print(f"  Wrong formulas sample: {wrong_formulas[:3]}")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {correct_count}/{total_data_rows} IF formulas correct")
            print(f"  Wrong formulas sample: {wrong_formulas[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Conditional formatting on F2:F45 with correct rules (0.25 points)
    # Rule 1 (priority 1): >90 -> red (#FF0000), Rule 2 (priority 2): >60 -> orange (#FFA500)
    try:
        cf_rules = ws.conditional_formatting
        cf_list = list(cf_rules)

        # Find CF range that covers F2:F45
        target_cf = None
        for cf in cf_list:
            cf_str = str(cf)
            if 'F2:F45' in cf_str or 'F2' in cf_str:
                target_cf = cf
                break

        if target_cf is None:
            print("FAIL: Component 3 — No conditional formatting found on F2:F45")
        else:
            rules = list(target_cf.rules)
            # Find red rule (>90) and orange rule (>60)
            red_rule_found = False
            orange_rule_found = False
            priority_correct = False

            red_rule_obj = None
            orange_rule_obj = None

            for rule in rules:
                if rule.type != 'cellIs':
                    continue
                # Check formula threshold
                formula_val = rule.formula[0] if rule.formula else None
                try:
                    threshold = float(formula_val)
                except (TypeError, ValueError):
                    continue

                # Check fill color
                fill_color = None
                if rule.dxf and rule.dxf.fill:
                    try:
                        fill_color = rule.dxf.fill.fgColor.rgb
                    except Exception:
                        pass

                if threshold == 90.0 and fill_color and fill_color.upper() in ('FFFF0000', 'FF0000'):
                    red_rule_found = True
                    red_rule_obj = rule
                elif threshold == 60.0 and fill_color and fill_color.upper() in ('FFFFA500', 'FFA500'):
                    orange_rule_found = True
                    orange_rule_obj = rule

            # Check priority: red (>90) must have higher priority (lower number) than orange (>60)
            if red_rule_obj and orange_rule_obj:
                if red_rule_obj.priority < orange_rule_obj.priority:
                    priority_correct = True

            if red_rule_found and orange_rule_found and priority_correct:
                print("PASS: Component 3 — CF rules correct: >90 red (FFFF0000, priority 1), >60 orange (FFFFA500, priority 2) (0.25 pts)")
                total_score += 0.25
            elif red_rule_found and orange_rule_found:
                # Both colors present but priority may be wrong
                partial = 0.15
                print(f"PARTIAL: Component 3 — Both CF rules found but priority order incorrect ({partial} pts)")
                print(f"  Red priority={red_rule_obj.priority if red_rule_obj else 'N/A'}, Orange priority={orange_rule_obj.priority if orange_rule_obj else 'N/A'}")
                total_score += partial
            elif red_rule_found or orange_rule_found:
                partial = 0.10
                print(f"PARTIAL: Component 3 — Only one CF rule found ({partial} pts)")
                print(f"  Red found: {red_rule_found}, Orange found: {orange_rule_found}")
                total_score += partial
            else:
                print("FAIL: Component 3 — Neither red nor orange CF rule found on F2:F45")
                if rules:
                    for r in rules:
                        print(f"  Rule: type={r.type}, formula={r.formula}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Row 1 frozen (freeze_panes == 'A2') (0.15 points)
    try:
        fp = ws.freeze_panes
        if fp == 'A2':
            print(f"PASS: Component 4 — freeze_panes='A2' (row 1 frozen) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — expected freeze_panes='A2', found: {repr(fp)}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
