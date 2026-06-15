"""
Initial Setup: Invoice Aging Report - Pre-task state
Task ID: calc_fin_invoice_aging_003
Domain: libreoffice_calc

Creates an invoice spreadsheet with 44 rows of realistic invoice data.
- Sheet 'Invoices' with headers: Invoice#, Client, Invoice Date, Amount, Status
- NO Days Outstanding column (F is empty)
- NO conditional formatting
- Row 1 NOT frozen
"""

import openpyxl
from openpyxl.styles import Font
from datetime import date, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_invoice_aging_003'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Invoices'

    # Headers
    headers = ['Invoice#', 'Client', 'Invoice Date', 'Amount', 'Status']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
        ws.cell(row=1, column=col).font = Font(bold=True)

    # Realistic invoice data (44 rows, rows 2-45)
    # Using dates that span 0-120 days ago from a reference of 2026-03-03
    # Mix of Paid and Unpaid to test IF formula
    clients = [
        'Apex Solutions Ltd', 'Bright Star Media', 'Cedar Grove Inc',
        'Delta Dynamics Corp', 'Emerald Technologies', 'Falcon Systems',
        'Global Trade Partners', 'Horizon Analytics', 'Ironclad Services',
        'JetSet Logistics', 'Keystone Ventures', 'Lighthouse Consulting',
        'Maple Leaf Designs', 'Nova Digital Agency', 'Offshore Capital Group',
        'Peak Performance LLC', 'Quantum Insights', 'Redwood Industries',
        'Silverline Financial', 'TechPath Solutions', 'Unified Networks',
        'Velocity Marketing', 'Westport Trading Co', 'Xenon Technologies',
        'Yellow Brick Media', 'Zenith Consulting', 'Altair Corp',
        'BlueSky Ventures', 'Clearwater Holdings', 'Dynamo Retail'
    ]

    invoice_data = [
        # (Invoice#, Client, Invoice Date, Amount, Status)
        # Days ago: mix of <60, 61-90, >90 for testing CF
        ('INV-2025-0091', 'Apex Solutions Ltd',       date(2025, 11, 1),  4250.00,  'Unpaid'),
        ('INV-2025-0092', 'Bright Star Media',        date(2025, 11, 8),  8750.50,  'Unpaid'),
        ('INV-2025-0093', 'Cedar Grove Inc',           date(2025, 11, 15), 1325.00,  'Paid'),
        ('INV-2025-0094', 'Delta Dynamics Corp',       date(2025, 11, 22), 6800.00,  'Unpaid'),
        ('INV-2025-0095', 'Emerald Technologies',      date(2025, 11, 30), 3150.75,  'Unpaid'),
        ('INV-2025-0096', 'Falcon Systems',            date(2025, 12, 5),  9400.00,  'Paid'),
        ('INV-2025-0097', 'Global Trade Partners',     date(2025, 12, 10), 2870.00,  'Unpaid'),
        ('INV-2025-0098', 'Horizon Analytics',         date(2025, 12, 15), 5300.00,  'Unpaid'),
        ('INV-2025-0099', 'Ironclad Services',         date(2025, 12, 20), 7625.50,  'Paid'),
        ('INV-2025-0100', 'JetSet Logistics',          date(2025, 12, 25), 1980.00,  'Unpaid'),
        ('INV-2026-0001', 'Keystone Ventures',         date(2026, 1, 2),   4430.00,  'Unpaid'),
        ('INV-2026-0002', 'Lighthouse Consulting',     date(2026, 1, 7),   3600.00,  'Paid'),
        ('INV-2026-0003', 'Maple Leaf Designs',        date(2026, 1, 10),  8250.00,  'Unpaid'),
        ('INV-2026-0004', 'Nova Digital Agency',       date(2026, 1, 14),  2140.50,  'Unpaid'),
        ('INV-2026-0005', 'Offshore Capital Group',    date(2026, 1, 18),  5875.00,  'Unpaid'),
        ('INV-2026-0006', 'Peak Performance LLC',      date(2026, 1, 22),  3320.00,  'Paid'),
        ('INV-2026-0007', 'Quantum Insights',          date(2026, 1, 25),  6100.00,  'Unpaid'),
        ('INV-2026-0008', 'Redwood Industries',        date(2026, 1, 28),  4780.75,  'Unpaid'),
        ('INV-2026-0009', 'Silverline Financial',      date(2026, 1, 31),  9950.00,  'Paid'),
        ('INV-2026-0010', 'TechPath Solutions',        date(2026, 2, 3),   1650.00,  'Unpaid'),
        ('INV-2026-0011', 'Unified Networks',          date(2026, 2, 6),   7230.00,  'Unpaid'),
        ('INV-2026-0012', 'Velocity Marketing',        date(2026, 2, 9),   3485.50,  'Unpaid'),
        ('INV-2026-0013', 'Westport Trading Co',       date(2026, 2, 11),  5120.00,  'Paid'),
        ('INV-2026-0014', 'Xenon Technologies',        date(2026, 2, 13),  2795.00,  'Unpaid'),
        ('INV-2026-0015', 'Yellow Brick Media',        date(2026, 2, 15),  4610.00,  'Unpaid'),
        ('INV-2026-0016', 'Zenith Consulting',         date(2026, 2, 17),  8340.00,  'Unpaid'),
        ('INV-2026-0017', 'Altair Corp',               date(2026, 2, 19),  1920.25,  'Paid'),
        ('INV-2026-0018', 'BlueSky Ventures',          date(2026, 2, 20),  6450.00,  'Unpaid'),
        ('INV-2026-0019', 'Clearwater Holdings',       date(2026, 2, 21),  3075.00,  'Unpaid'),
        ('INV-2026-0020', 'Dynamo Retail',             date(2026, 2, 22),  5890.50,  'Paid'),
        ('INV-2026-0021', 'Apex Solutions Ltd',        date(2026, 2, 23),  2250.00,  'Unpaid'),
        ('INV-2026-0022', 'Bright Star Media',         date(2026, 2, 24),  7180.00,  'Unpaid'),
        ('INV-2026-0023', 'Cedar Grove Inc',            date(2026, 2, 25),  3940.00,  'Unpaid'),
        ('INV-2026-0024', 'Delta Dynamics Corp',        date(2026, 2, 25),  4560.00,  'Paid'),
        ('INV-2026-0025', 'Emerald Technologies',       date(2026, 2, 26),  6720.75,  'Unpaid'),
        ('INV-2026-0026', 'Falcon Systems',             date(2026, 2, 26),  1840.00,  'Unpaid'),
        ('INV-2026-0027', 'Global Trade Partners',      date(2026, 2, 27),  9100.00,  'Unpaid'),
        ('INV-2026-0028', 'Horizon Analytics',          date(2026, 2, 27),  3350.00,  'Paid'),
        ('INV-2026-0029', 'Ironclad Services',          date(2026, 2, 28),  5630.00,  'Unpaid'),
        ('INV-2026-0030', 'JetSet Logistics',           date(2026, 2, 28),  2490.50,  'Unpaid'),
        ('INV-2026-0031', 'Keystone Ventures',          date(2026, 3, 1),   7860.00,  'Unpaid'),
        ('INV-2026-0032', 'Lighthouse Consulting',      date(2026, 3, 1),   4175.00,  'Paid'),
        ('INV-2026-0033', 'Maple Leaf Designs',         date(2026, 3, 2),   3010.00,  'Unpaid'),
        ('INV-2026-0034', 'Nova Digital Agency',        date(2026, 3, 2),   6540.00,  'Unpaid'),
    ]

    for r, (inv_num, client, inv_date, amount, status) in enumerate(invoice_data, 2):
        ws.cell(row=r, column=1, value=inv_num)
        ws.cell(row=r, column=2, value=client)
        ws.cell(row=r, column=3, value=inv_date)
        ws.cell(row=r, column=3).number_format = 'yyyy-mm-dd'
        ws.cell(row=r, column=4, value=amount)
        ws.cell(row=r, column=4).number_format = '#,##0.00'
        ws.cell(row=r, column=5, value=status)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10

    # NO freeze panes (task requires adding it)
    # NO column F / Days Outstanding (task requires adding it)
    # NO conditional formatting (task requires adding it)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheets: {wb.sheetnames}')
    print(f'  Rows (data): {len(invoice_data)}')

create_initial()
