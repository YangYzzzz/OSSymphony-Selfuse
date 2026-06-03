"""
Initial Setup: Insert hyperlink in cell A1 of Links sheet
Task ID: calc_gg1_011
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_011'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Links (the task target sheet - must be empty) ---
    ws_links = wb.active
    ws_links.title = 'Links'
    # All cells blank - this is where the agent will add the hyperlink

    # Set column width for readability when opened
    ws_links.column_dimensions['A'].width = 35
    ws_links.column_dimensions['B'].width = 45

    # Add headers in row 1 of a different area to provide context
    # Actually, per task: "The 'Links' sheet is currently empty - all cells are blank"
    # So we leave it completely empty.

    # --- Sheet 2: Resources (additional content for realism) ---
    ws_res = wb.create_sheet('Resources')
    headers = ['Category', 'Description', 'Last Updated', 'Priority']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws_res.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align

    data = [
        ['Office Suite', 'Documentation and templates for team use', '2025-11-20', 'High'],
        ['Design Tools', 'Graphics and UI mockup resources', '2025-10-15', 'Medium'],
        ['Development', 'Code repositories and API references', '2025-12-01', 'High'],
        ['Communication', 'Team messaging and video tools', '2025-09-30', 'Medium'],
        ['Project Management', 'Task tracking and sprint planning', '2025-11-05', 'High'],
        ['Cloud Storage', 'File sharing and backup services', '2025-08-22', 'Low'],
        ['Analytics', 'Data visualization and reporting tools', '2025-10-28', 'Medium'],
        ['Security', 'VPN, password managers, and compliance', '2025-12-10', 'High'],
        ['Training', 'Online courses and certification platforms', '2025-07-14', 'Low'],
        ['HR & Admin', 'Payroll, benefits, and onboarding portals', '2025-11-18', 'Medium'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws_res.cell(row=r, column=c, value=val)

    ws_res.column_dimensions['A'].width = 22
    ws_res.column_dimensions['B'].width = 45
    ws_res.column_dimensions['C'].width = 16
    ws_res.column_dimensions['D'].width = 12

    # --- Sheet 3: Team Members ---
    ws_team = wb.create_sheet('Team')
    team_headers = ['Name', 'Role', 'Department', 'Email']
    for col, h in enumerate(team_headers, 1):
        cell = ws_team.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="FF548235", end_color="FF548235", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    team_data = [
        ['Sarah Chen', 'Project Lead', 'Engineering', 'sarah.chen@company.com'],
        ['Marcus Johnson', 'UX Designer', 'Design', 'marcus.j@company.com'],
        ['Elena Rodriguez', 'Backend Dev', 'Engineering', 'elena.r@company.com'],
        ['David Kim', 'Data Analyst', 'Analytics', 'david.kim@company.com'],
        ['Priya Patel', 'QA Engineer', 'Engineering', 'priya.p@company.com'],
        ['James Wright', 'Scrum Master', 'PMO', 'james.w@company.com'],
        ['Aisha Mohammed', 'Frontend Dev', 'Engineering', 'aisha.m@company.com'],
        ['Lucas Weber', 'DevOps Lead', 'Infrastructure', 'lucas.w@company.com'],
    ]

    for r, row_data in enumerate(team_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_team.cell(row=r, column=c, value=val)

    ws_team.column_dimensions['A'].width = 20
    ws_team.column_dimensions['B'].width = 18
    ws_team.column_dimensions['C'].width = 18
    ws_team.column_dimensions['D'].width = 30

    # Ensure Links is the active sheet (first sheet)
    wb.active = wb.sheetnames.index('Links')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
