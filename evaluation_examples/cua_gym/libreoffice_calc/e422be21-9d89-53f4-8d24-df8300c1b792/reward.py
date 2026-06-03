"""
Reward Script: Insert hyperlink in cell A1 of 'Links' sheet
Task ID: calc_gg1_011
Domain: libreoffice_calc
Scoring:
  Component 1: Cell A1 has a hyperlink object (0.3 pts)
  Component 2: Hyperlink target URL is 'https://www.libreoffice.org' (0.3 pts)
  Component 3: Cell A1 display text is 'LibreOffice Official Site' (0.25 pts)
  Component 4: Font styled as hyperlink (blue + underline) (0.15 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_011'


def persist_app_state(domain):
    """Best-effort save of any open LibreOffice document."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify that a hyperlink was inserted in cell A1 of the 'Links' sheet
    with URL 'https://www.libreoffice.org' and display text 'LibreOffice Official Site'.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Links' sheet must exist
    if 'Links' not in wb.sheetnames:
        print("FAIL: 'Links' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Links']
    cell = ws['A1']

    # Component 1: Cell A1 has a hyperlink object (0.3 points)
    try:
        if cell.hyperlink is not None:
            print(f"PASS: Component 1 — A1 has a hyperlink object (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — A1 has no hyperlink (expected a hyperlink object)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Hyperlink target URL is 'https://www.libreoffice.org' (0.3 points)
    try:
        if cell.hyperlink is not None and cell.hyperlink.target is not None:
            target = cell.hyperlink.target.rstrip('/')
            expected = 'https://www.libreoffice.org'
            if target == expected:
                print(f"PASS: Component 2 — Hyperlink target is '{target}' (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — Hyperlink target is '{target}', expected '{expected}'")
        else:
            print(f"FAIL: Component 2 — No hyperlink or no target URL on A1")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Cell A1 display text is 'LibreOffice Official Site' (0.25 points)
    try:
        val = cell.value
        if val is not None and str(val).strip() == 'LibreOffice Official Site':
            print(f"PASS: Component 3 — A1 display text is '{val}' (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 3 — A1 value is {repr(val)}, expected 'LibreOffice Official Site'")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Font styled as hyperlink — blue color and underline (0.15 points)
    try:
        font = cell.font
        has_underline = font.underline is not None and font.underline != 'none'
        # Check for blue-ish color (common hyperlink colors)
        has_blue = False
        if font.color and font.color.rgb:
            rgb = font.color.rgb
            # Accept any blue-dominated color: last 2 hex chars should be high
            # Standard blue hyperlink colors: 0000FF, 0563C1, 0000EE, etc.
            if isinstance(rgb, str) and len(rgb) >= 6:
                # Take last 6 chars (strip alpha if 8-char ARGB)
                hex_rgb = rgb[-6:]
                r_val = int(hex_rgb[0:2], 16)
                b_val = int(hex_rgb[4:6], 16)
                if b_val > 128 and b_val > r_val:
                    has_blue = True

        if has_underline and has_blue:
            print(f"PASS: Component 4 — Font has underline='{font.underline}' and blue color rgb='{font.color.rgb}' (0.15 pts)")
            total_score += 0.15
        elif has_underline:
            print(f"PARTIAL: Component 4 — Font has underline but color is not blue (rgb={font.color.rgb if font.color else 'None'})")
            total_score += 0.075
        elif has_blue:
            print(f"PARTIAL: Component 4 — Font has blue color but no underline")
            total_score += 0.075
        else:
            print(f"FAIL: Component 4 — Font not styled as hyperlink (underline={font.underline}, color={font.color.rgb if font.color else 'None'})")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist any unsaved GUI state before verification
persist_app_state("libreoffice_calc")

# Test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
