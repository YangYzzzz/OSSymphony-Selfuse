"""
Initial Setup: Create attainment spreadsheet with sales rep quota data
Task ID: calc_sales_043
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_043'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attainment'

    # Headers
    ws['A1'] = 'Sales Rep'
    ws['B1'] = 'Attainment %'

    # Style headers
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal='center')
    for cell in [ws['A1'], ws['B1']]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")

    # Data rows
    data = [
        ['Alice', 1.15],
        ['Bob', 0.82],
        ['Carol', 1.05],
        ['Dan', 0.65],
        ['Eve', 0.98],
        ['Frank', 1.30],
        ['Grace', 0.75],
    ]

    for r, (name, pct) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=pct)
        ws.cell(row=r, column=2).number_format = '0%'

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 16

    # NO conditional formatting in initial state

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
