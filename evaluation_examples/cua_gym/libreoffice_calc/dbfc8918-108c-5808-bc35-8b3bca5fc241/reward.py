"""
Reward Script: Make a pivot table from expense report data showing total expenses by category
Task ID: calc_pivot_005
Domain: libreoffice_calc
Scoring:
  Component 1 (0.2): Pivot/summary sheet exists (not in initial)
  Component 2 (0.2): All 5 category labels present
  Component 3 (0.4): Category amounts match expected values
  Component 4 (0.2): Grand total is 54000
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_005'

# Expected pivot table values from task context
EXPECTED_CATEGORIES = {
    'Travel': 18500,
    'Meals': 8200,
    'Supplies': 5400,
    'Software': 12300,
    'Training': 9600,
}
EXPECTED_GRAND_TOTAL = 54000


def find_pivot_sheet(wb):
    """Find the pivot/summary sheet (any sheet other than 'Expenses')."""
    for name in wb.sheetnames:
        if name.lower() != 'expenses':
            return wb[name]
    return None


def read_pivot_data(ws):
    """
    Read the pivot sheet and extract category->amount mapping and grand total.
    Searches all rows for category names and numeric values.
    Returns (category_amounts_dict, grand_total_or_None).
    """
    categories_found = {}
    grand_total = None

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        label = None
        amount = None
        for cell in row:
            val = cell.value
            if val is None:
                continue
            if isinstance(val, str):
                stripped = val.strip()
                # Check for category names
                for cat in EXPECTED_CATEGORIES:
                    if stripped.lower() == cat.lower():
                        label = cat
                        break
                # Check for grand total label
                if 'grand' in stripped.lower() and 'total' in stripped.lower():
                    label = 'Grand Total'
            elif isinstance(val, (int, float)):
                amount = val

        if label and amount is not None:
            if label == 'Grand Total':
                grand_total = amount
            else:
                categories_found[label] = amount

    return categories_found, grand_total


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: A pivot/summary sheet exists beyond the original 'Expenses' sheet (0.2 points)
    # This FAILS on initial (only 'Expenses') and PASSES on golden (has 'Pivot')
    try:
        pivot_ws = find_pivot_sheet(wb)
        if pivot_ws is not None:
            print(f"PASS: Component 1 — Pivot sheet found: '{pivot_ws.title}' (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 1 — No pivot/summary sheet found (only sheets: {wb.sheetnames})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    if pivot_ws is None:
        # No pivot sheet means no further checks possible
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    # Read all pivot data
    try:
        cat_amounts, grand_total = read_pivot_data(pivot_ws)
        print(f"  Parsed categories: {cat_amounts}")
        print(f"  Parsed grand total: {grand_total}")
    except Exception as e:
        print(f"ERROR: Could not parse pivot data: {e}")
        cat_amounts, grand_total = {}, None

    # Component 2: All 5 category labels are present in the pivot sheet (0.2 points)
    # This FAILS on initial (no pivot sheet) and PASSES on golden
    try:
        found_cats = set(cat_amounts.keys())
        expected_cats = set(EXPECTED_CATEGORIES.keys())
        if expected_cats.issubset(found_cats):
            print(f"PASS: Component 2 — All 5 categories present: {sorted(found_cats)} (0.2 pts)")
            total_score += 0.2
        else:
            missing = expected_cats - found_cats
            print(f"FAIL: Component 2 — Missing categories: {missing} (found: {sorted(found_cats)})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Category amounts match expected values (0.4 points)
    # Each correct category contributes 0.08 points (5 * 0.08 = 0.4)
    # This FAILS on initial (no pivot sheet) and PASSES on golden
    try:
        correct_count = 0
        for cat, expected_amt in EXPECTED_CATEGORIES.items():
            actual_amt = cat_amounts.get(cat)
            if actual_amt is not None and abs(float(actual_amt) - expected_amt) <= 1.0:
                print(f"  PASS: {cat} = {actual_amt} (expected {expected_amt})")
                correct_count += 1
                total_score += 0.08
            else:
                print(f"  FAIL: {cat} = {actual_amt} (expected {expected_amt})")
        print(f"Component 3: {correct_count}/5 categories correct ({correct_count * 0.08:.2f} pts)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Grand total is 54000 (0.2 points)
    # This FAILS on initial (no pivot sheet) and PASSES on golden
    try:
        if grand_total is not None and abs(float(grand_total) - EXPECTED_GRAND_TOTAL) <= 1.0:
            print(f"PASS: Component 4 — Grand Total = {grand_total} (expected {EXPECTED_GRAND_TOTAL}) (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 4 — Grand Total = {grand_total} (expected {EXPECTED_GRAND_TOTAL})")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
