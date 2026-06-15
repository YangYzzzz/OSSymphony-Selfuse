"""
Initial Setup: Create expense report data for pivot table task
Task ID: calc_pivot_005
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_005'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"

    # Headers
    headers = ['ExpenseID', 'Date', 'Employee', 'Category', 'Amount', 'Approved']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12

    # Employees
    employees = [
        'Sarah Chen', 'Marcus Johnson', 'Emily Rodriguez', 'David Kim',
        'Jessica Williams', 'Michael Brown', 'Amanda Garcia', 'Robert Taylor',
        'Lisa Anderson', 'James Wilson', 'Rachel Martinez', 'Daniel Lee',
        'Sophia Thomas', 'Christopher White', 'Olivia Harris', 'Matthew Clark',
        'Emma Lewis', 'Andrew Robinson', 'Natalie Walker', 'Kevin Hall',
    ]

    categories = ['Travel', 'Meals', 'Supplies', 'Software', 'Training']

    # Target totals per category (must match ground truth exactly)
    # Travel=18500, Meals=8200, Supplies=5400, Software=12300, Training=9600
    # Total=54000, 120 rows

    # Distribute rows: Travel 30, Meals 25, Supplies 20, Software 25, Training 20 = 120
    category_targets = {
        'Travel':   {'count': 30, 'total': 18500},
        'Meals':    {'count': 25, 'total': 8200},
        'Supplies': {'count': 20, 'total': 5400},
        'Software': {'count': 25, 'total': 12300},
        'Training': {'count': 20, 'total': 9600},
    }

    # Generate amounts for each category that sum to the target
    rows = []
    for cat, info in category_targets.items():
        count = info['count']
        total = info['total']
        # Generate count-1 random amounts, then compute last to hit exact total
        if count == 1:
            amounts = [total]
        else:
            avg = total / count
            raw = []
            for i in range(count - 1):
                # Random variation around average
                lo = max(20, avg * 0.3)
                hi = avg * 1.8
                raw.append(round(random.uniform(lo, hi), 2))
            current_sum = sum(raw)
            last = round(total - current_sum, 2)
            # Ensure last is positive
            if last < 10:
                # Redistribute: reduce the largest
                deficit = 10 - last
                raw[0] -= (deficit + 5)
                last = round(total - sum(raw), 2)
            raw.append(last)
            amounts = raw

        for amt in amounts:
            rows.append({
                'category': cat,
                'amount': amt,
            })

    # Shuffle rows so categories are mixed
    random.shuffle(rows)

    # 2024 dates spread across the year
    months_days = []
    for m in range(1, 13):
        for d in [3, 7, 12, 15, 18, 21, 24, 27, 29]:
            if m == 2 and d == 29:
                continue
            months_days.append(f'2024-{m:02d}-{d:02d}')

    for i, row_data in enumerate(rows):
        r = i + 2  # row (1-indexed, header is row 1)
        expense_id = i + 1
        date_str = months_days[i % len(months_days)]
        employee = employees[i % len(employees)]
        category = row_data['category']
        amount = row_data['amount']
        approved = random.choice(['Yes', 'Yes', 'Yes', 'No'])  # ~75% approved

        ws.cell(row=r, column=1, value=expense_id)
        ws.cell(row=r, column=2, value=date_str)
        ws.cell(row=r, column=3, value=employee)
        ws.cell(row=r, column=4, value=category)
        amt_cell = ws.cell(row=r, column=5, value=amount)
        amt_cell.number_format = '#,##0.00'
        ws.cell(row=r, column=6, value=approved)

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Verify totals
    totals = {}
    for i in range(len(rows)):
        cat = rows[i]['category']
        amt = rows[i]['amount']
        totals[cat] = totals.get(cat, 0) + amt
    for cat, total in sorted(totals.items()):
        print(f'  {cat}: {total:.2f}')
    print(f'  Grand Total: {sum(totals.values()):.2f}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
