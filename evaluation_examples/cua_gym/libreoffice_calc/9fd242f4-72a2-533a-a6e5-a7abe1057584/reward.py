"""
Reward Script: Format expense report with currency, total row, and bold headers/totals
Task ID: calc_fin_expense_currency_002
Domain: libreoffice_calc
Scoring:
  - Component 1: Currency format ([$$-409]#,##0.00) applied to D2:F30 (0.35 pts)
  - Component 2: Total row at row 32 with SUM formulas and currency format (0.35 pts)
  - Component 3: Header row 1 (A1:F1) is bold (0.15 pts)
  - Component 4: Total row 32 (A32:F32) is bold (0.15 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_expense_currency_002'
CURRENCY_FORMAT = '[$$-409]#,##0.00'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists
    if 'Expenses' not in wb.sheetnames:
        print("CRITICAL: 'Expenses' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Expenses']

    # Component 1: Currency format applied to D2:D30, E2:E30, F2:F30 (0.35 points)
    # Task requires: D2:D30, E2:E30, F2:F30 formatted as [$$-409]#,##0.00
    # This FAILS on initial (all 'General') and PASSES on golden
    try:
        currency_cols = [4, 5, 6]  # D, E, F
        formatted_count = 0
        total_cells = 29 * 3  # rows 2-30, cols D/E/F = 87 cells

        for col in currency_cols:
            for row in range(2, 31):
                cell = ws.cell(row=row, column=col)
                if cell.number_format == CURRENCY_FORMAT:
                    formatted_count += 1

        if formatted_count == total_cells:
            print(f"PASS: Component 1 — All {total_cells} cost cells (D2:F30) have currency format (0.35 pts)")
            total_score += 0.35
        elif formatted_count > 0:
            ratio = formatted_count / total_cells
            partial = round(0.35 * ratio, 4)
            print(f"PARTIAL: Component 1 — {formatted_count}/{total_cells} cells have currency format "
                  f"(partial {partial} pts)")
            # Only award full 0.35 for full completion — no partial here per task requirement
            print(f"FAIL: Component 1 — expected all {total_cells} currency cells, got {formatted_count}")
        else:
            print(f"FAIL: Component 1 — No cells in D2:F30 have currency format (found 0/{total_cells})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Total row at row 32 with SUM formulas + currency format (0.35 points)
    # Task requires: A32='Total', D32=SUM(D2:D30), E32=SUM(E2:E30), F32=SUM(F2:F30), D32:F32 currency
    # This FAILS on initial (row 32 empty) and PASSES on golden
    try:
        sum_formula_count = 0
        sum_currency_count = 0

        # Check A32 label
        a32_val = ws.cell(row=32, column=1).value
        label_ok = (a32_val is not None and str(a32_val).strip().lower() == 'total')

        # Check D32, E32, F32 for SUM formulas and currency format
        expected_formulas = {
            4: '=SUM(D2:D30)',
            5: '=SUM(E2:E30)',
            6: '=SUM(F2:F30)',
        }
        for col, expected_formula in expected_formulas.items():
            cell = ws.cell(row=32, column=col)
            val = cell.value
            # Check formula (case-insensitive, whitespace-insensitive)
            if isinstance(val, str):
                val_norm = val.upper().replace(' ', '')
                expected_norm = expected_formula.upper().replace(' ', '')
                if val_norm == expected_norm:
                    sum_formula_count += 1
            # Check currency format
            if cell.number_format == CURRENCY_FORMAT:
                sum_currency_count += 1

        if label_ok and sum_formula_count == 3 and sum_currency_count == 3:
            print(f"PASS: Component 2 — Row 32 total row: label='Total', "
                  f"SUM formulas in D32/E32/F32, all currency-formatted (0.35 pts)")
            total_score += 0.35
        else:
            details = []
            if not label_ok:
                details.append(f"A32 label is {repr(a32_val)}, expected 'Total'")
            if sum_formula_count < 3:
                details.append(f"SUM formulas: {sum_formula_count}/3 correct")
            if sum_currency_count < 3:
                details.append(f"Currency format on total row: {sum_currency_count}/3 cells")
            print(f"FAIL: Component 2 — {'; '.join(details)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Header row 1 (A1:F1) is bold (0.15 points)
    # Task requires: bold headers in row 1
    # This FAILS on initial (all non-bold) and PASSES on golden
    try:
        bold_header_count = 0
        for col in range(1, 7):  # A through F
            cell = ws.cell(row=1, column=col)
            if cell.font.bold:
                bold_header_count += 1

        if bold_header_count == 6:
            print(f"PASS: Component 3 — All 6 header cells (A1:F1) are bold (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 — Only {bold_header_count}/6 header cells (A1:F1) are bold")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Total row 32 (A32:F32) is bold (0.15 points)
    # Task requires: bold total row
    # This FAILS on initial (row 32 absent/not bold) and PASSES on golden
    try:
        bold_total_count = 0
        # Check the cells that should have content/formatting: A32 and D32:F32
        # B32 and C32 may or may not be bold depending on implementation
        # We check A32, D32, E32, F32 which are the meaningful cells
        key_cols = [1, 4, 5, 6]  # A, D, E, F
        for col in key_cols:
            cell = ws.cell(row=32, column=col)
            if cell.font.bold:
                bold_total_count += 1

        if bold_total_count == len(key_cols):
            print(f"PASS: Component 4 — Total row cells (A32, D32:F32) are bold (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — Only {bold_total_count}/{len(key_cols)} total row cells are bold "
                  f"(A32, D32, E32, F32)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
