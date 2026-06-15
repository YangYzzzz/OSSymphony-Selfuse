"""
Initial Setup: Expense report with plain number amounts (no currency formatting)
Task ID: calc_fin_expense_currency_002
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_expense_currency_002'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Expenses'

    # --- Row 1: Headers (NOT bold, plain style) ---
    headers = ['Date', 'Employee', 'Category', 'Amount', 'Tax', 'Total']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # --- Rows 2-30: 29 expense entries (plain numbers, no currency format) ---
    data = [
        ('2025-01-05', 'Sarah Chen',       'Travel',        412.50,  37.13,  449.63),
        ('2025-01-07', 'Marcus Johnson',   'Office Supplies', 89.99,   8.10,   98.09),
        ('2025-01-10', 'Priya Patel',      'Meals',         134.80,  12.13,  146.93),
        ('2025-01-12', 'Derek Williams',   'Software',     1250.00, 112.50, 1362.50),
        ('2025-01-15', 'Aisha Okonkwo',    'Travel',        875.20,  78.77,  953.97),
        ('2025-01-18', 'Tom Nguyen',       'Equipment',    3200.00, 288.00, 3488.00),
        ('2025-01-20', 'Sarah Chen',       'Meals',          67.45,   6.07,   73.52),
        ('2025-01-22', 'Julia Rossi',      'Office Supplies', 145.30,  13.08,  158.38),
        ('2025-01-25', 'Marcus Johnson',   'Travel',        520.00,  46.80,  566.80),
        ('2025-01-28', 'Priya Patel',      'Software',      599.00,  53.91,  652.91),
        ('2025-02-02', 'Derek Williams',   'Meals',          92.60,   8.33,  100.93),
        ('2025-02-05', 'Aisha Okonkwo',    'Equipment',    1850.00, 166.50, 2016.50),
        ('2025-02-07', 'Tom Nguyen',       'Travel',        340.75,  30.67,  371.42),
        ('2025-02-10', 'Sarah Chen',       'Office Supplies', 55.20,   4.97,   60.17),
        ('2025-02-12', 'Julia Rossi',      'Software',      299.99,  27.00,  326.99),
        ('2025-02-14', 'Marcus Johnson',   'Equipment',    2750.00, 247.50, 2997.50),
        ('2025-02-17', 'Priya Patel',      'Travel',        680.00,  61.20,  741.20),
        ('2025-02-20', 'Derek Williams',   'Meals',         118.40,  10.66,  129.06),
        ('2025-02-22', 'Aisha Okonkwo',    'Office Supplies', 210.75,  18.97,  229.72),
        ('2025-02-25', 'Tom Nguyen',       'Software',      450.00,  40.50,  490.50),
        ('2025-03-01', 'Sarah Chen',       'Equipment',    4100.00, 369.00, 4469.00),
        ('2025-03-04', 'Julia Rossi',      'Travel',        295.60,  26.60,  322.20),
        ('2025-03-06', 'Marcus Johnson',   'Meals',          75.90,   6.83,   82.73),
        ('2025-03-09', 'Priya Patel',      'Office Supplies', 180.00,  16.20,  196.20),
        ('2025-03-12', 'Derek Williams',   'Software',      899.00,  80.91,  979.91),
        ('2025-03-15', 'Aisha Okonkwo',    'Travel',       1125.00, 101.25, 1226.25),
        ('2025-03-18', 'Tom Nguyen',       'Meals',          88.20,   7.94,   96.14),
        ('2025-03-20', 'Sarah Chen',       'Equipment',    2300.00, 207.00, 2507.00),
        ('2025-03-24', 'Julia Rossi',      'Office Supplies', 62.50,   5.63,   68.13),
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])  # Date as string
        ws.cell(row=r, column=2, value=row_data[1])  # Employee
        ws.cell(row=r, column=3, value=row_data[2])  # Category
        ws.cell(row=r, column=4, value=row_data[3])  # Amount (plain number)
        ws.cell(row=r, column=5, value=row_data[4])  # Tax (plain number)
        ws.cell(row=r, column=6, value=row_data[5])  # Total (plain number)

    # Row 31 is left empty (as specified in context)
    # No row 32 (no total row yet)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Expenses')
    print(f'  Rows: 1 header + 29 data rows (rows 2-30)')
    print(f'  Row 31: empty')
    print(f'  Columns D, E, F: plain numbers (no currency format)')
    print(f'  Headers: NOT bold')
    print(f'  No total row')


create_initial()
