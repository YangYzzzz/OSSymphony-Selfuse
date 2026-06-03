"""
Reward Script: Build a year-over-year comparison table in Sheet2
Task ID: osworld_calc_sheet2_summary_table_007
Domain: libreoffice_calc

Scoring Rubric (total = 1.0):
  Component 1 (0.25): Summary sheet has a table structure — at least 3 rows and 7 columns of data
  Component 2 (0.25): Header row contains all required column labels
  Component 3 (0.25): All 5 expected departments appear as row labels in the table
  Component 4 (0.25): Data cells contain COUNTIFS, AVERAGEIFS, SUMIFS, and % change formulas
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_sheet2_summary_table_007'

# Expected departments (from HR Data sheet)
EXPECTED_DEPARTMENTS = {'Engineering', 'Finance', 'HR', 'Marketing', 'Operations'}

# Required header keywords (case-insensitive partial match)
REQUIRED_HEADER_KEYWORDS = [
    '2023',
    '2024',
    'headcount',
    'salary',
    'payroll',
    'change',
]


def normalize(val):
    """Normalize a cell value to a lowercase stripped string."""
    if val is None:
        return ''
    return str(val).strip().lower()


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook — fail fast if the file can't be opened
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: "Summary" sheet must exist
    if 'Summary' not in wb.sheetnames:
        print("FAIL: 'Summary' sheet not found in workbook")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws = wb['Summary']

    # ------------------------------------------------------------------
    # Component 1: Table structure — at least 3 rows and 7 columns (0.25)
    # Initial env: Summary sheet is empty (max_row=1, max_column=1, value=None)
    # Golden env:  Summary has 7 rows and 8 columns populated
    # ------------------------------------------------------------------
    try:
        # Collect all non-empty cells to gauge actual data extent
        non_empty_rows = set()
        non_empty_cols = set()
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and str(cell.value).strip() != '':
                    non_empty_rows.add(cell.row)
                    non_empty_cols.add(cell.column)

        has_min_rows = len(non_empty_rows) >= 3   # header row + at least 2 dept rows
        has_min_cols = len(non_empty_cols) >= 7   # 7+ columns of data

        if has_min_rows and has_min_cols:
            print(f"PASS: Component 1 — table structure present "
                  f"({len(non_empty_rows)} rows, {len(non_empty_cols)} cols) (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — insufficient table structure "
                  f"(found {len(non_empty_rows)} non-empty rows, "
                  f"{len(non_empty_cols)} non-empty cols; need >= 3 rows and >= 7 cols)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ------------------------------------------------------------------
    # Component 2: Header row contains required column labels (0.25)
    # We look for a row that contains all required keywords distributed across cells.
    # Initial env: Summary is empty → no header row → FAIL
    # ------------------------------------------------------------------
    try:
        matching_header_rows = [
            row for row in ws.iter_rows()
            if all(kw in ' '.join(normalize(cell.value) for cell in row)
                   for kw in REQUIRED_HEADER_KEYWORDS)
        ]
        if len(matching_header_rows) >= 1:
            header_row_values = [cell.value for cell in matching_header_rows[0]]
            print(f"PASS: Component 2 — header row found: {header_row_values} (0.25 pts)")
            total_score += 0.25
        else:
            print("FAIL: Component 2 — no single row contains all required header keywords: "
                  f"{REQUIRED_HEADER_KEYWORDS}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ------------------------------------------------------------------
    # Component 3: All 5 expected departments are listed as row labels (0.25)
    # Initial env: Summary is empty → no departments → FAIL
    # ------------------------------------------------------------------
    try:
        # Collect all string values in column A (or scan all cells)
        found_departments = set()
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    dept = cell.value.strip()
                    if dept in EXPECTED_DEPARTMENTS:
                        found_departments.add(dept)

        if found_departments == EXPECTED_DEPARTMENTS:
            print(f"PASS: Component 3 — all 5 departments present: "
                  f"{sorted(found_departments)} (0.25 pts)")
            total_score += 0.25
        else:
            missing = EXPECTED_DEPARTMENTS - found_departments
            print(f"FAIL: Component 3 — missing departments: {sorted(missing)}; "
                  f"found: {sorted(found_departments)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ------------------------------------------------------------------
    # Component 4: Data cells use COUNTIFS, AVERAGEIFS, SUMIFS, and % change formulas (0.25)
    # Task requires formulas computed from Sheet1 (HR Data) using these functions.
    # Initial env: Summary is empty → no formulas → FAIL
    # ------------------------------------------------------------------
    try:
        formula_flags = {
            'COUNTIFS': False,
            'AVERAGEIFS': False,
            'SUMIFS': False,
            'percent_change': False,   # any formula involving subtraction/division for % change
        }

        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    upper_val = cell.value.upper()
                    if 'COUNTIFS' in upper_val:
                        formula_flags['COUNTIFS'] = True
                    if 'AVERAGEIFS' in upper_val:
                        formula_flags['AVERAGEIFS'] = True
                    if 'SUMIFS' in upper_val:
                        formula_flags['SUMIFS'] = True
                    # % change formula pattern: involves division and subtraction of two payroll cells
                    # e.g., =(G3-D3)/D3
                    if ('/' in upper_val and '-' in upper_val
                            and 'SUMIFS' not in upper_val
                            and 'AVERAGEIFS' not in upper_val
                            and 'COUNTIFS' not in upper_val):
                        formula_flags['percent_change'] = True

        all_formulas_present = all(formula_flags.values())
        if all_formulas_present:
            print(f"PASS: Component 4 — all required formula types found "
                  f"(COUNTIFS, AVERAGEIFS, SUMIFS, % change) (0.25 pts)")
            total_score += 0.25
        else:
            missing_formulas = [k for k, v in formula_flags.items() if not v]
            print(f"FAIL: Component 4 — missing formula types: {missing_formulas}; "
                  f"status: {formula_flags}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point: test against canonical artifact path on the VM
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
