"""
Initial Setup: Year-over-year HR comparison table task
Task ID: osworld_calc_sheet2_summary_table_007
Domain: libreoffice_calc

Creates a spreadsheet with HR records on Sheet1 spanning 2023-2024.
Sheet2 is intentionally left empty — the agent must build the summary table.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_sheet2_summary_table_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ----------------------------------------------------------------
    # Sheet1: HR Data (Year, Department, Employee ID, Annual Salary)
    # ----------------------------------------------------------------
    ws1 = wb.active
    ws1.title = 'HR Data'

    # Headers
    headers = ['Year', 'Department', 'Employee ID', 'Annual Salary']
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Realistic HR data: multiple departments, 2023 and 2024
    data = [
        # Year, Department, Employee ID, Annual Salary
        # 2023 Engineering
        [2023, 'Engineering', 'EMP-1001', 92000],
        [2023, 'Engineering', 'EMP-1002', 87500],
        [2023, 'Engineering', 'EMP-1003', 105000],
        [2023, 'Engineering', 'EMP-1004', 78000],
        [2023, 'Engineering', 'EMP-1005', 115000],
        # 2023 Marketing
        [2023, 'Marketing', 'EMP-2001', 68000],
        [2023, 'Marketing', 'EMP-2002', 74500],
        [2023, 'Marketing', 'EMP-2003', 71000],
        [2023, 'Marketing', 'EMP-2004', 62000],
        # 2023 Finance
        [2023, 'Finance', 'EMP-3001', 95000],
        [2023, 'Finance', 'EMP-3002', 88000],
        [2023, 'Finance', 'EMP-3003', 102000],
        # 2023 HR
        [2023, 'HR', 'EMP-4001', 65000],
        [2023, 'HR', 'EMP-4002', 60000],
        [2023, 'HR', 'EMP-4003', 72000],
        # 2023 Operations
        [2023, 'Operations', 'EMP-5001', 58000],
        [2023, 'Operations', 'EMP-5002', 61500],
        [2023, 'Operations', 'EMP-5003', 55000],
        [2023, 'Operations', 'EMP-5004', 63000],
        # 2024 Engineering
        [2024, 'Engineering', 'EMP-1001', 96000],
        [2024, 'Engineering', 'EMP-1002', 91000],
        [2024, 'Engineering', 'EMP-1003', 110000],
        [2024, 'Engineering', 'EMP-1004', 82000],
        [2024, 'Engineering', 'EMP-1005', 120000],
        [2024, 'Engineering', 'EMP-1006', 89000],
        # 2024 Marketing
        [2024, 'Marketing', 'EMP-2001', 72000],
        [2024, 'Marketing', 'EMP-2002', 78000],
        [2024, 'Marketing', 'EMP-2003', 74500],
        [2024, 'Marketing', 'EMP-2004', 65000],
        [2024, 'Marketing', 'EMP-2005', 69000],
        # 2024 Finance
        [2024, 'Finance', 'EMP-3001', 100000],
        [2024, 'Finance', 'EMP-3002', 93000],
        [2024, 'Finance', 'EMP-3003', 108000],
        [2024, 'Finance', 'EMP-3004', 85000],
        # 2024 HR
        [2024, 'HR', 'EMP-4001', 68000],
        [2024, 'HR', 'EMP-4002', 63500],
        [2024, 'HR', 'EMP-4003', 76000],
        # 2024 Operations
        [2024, 'Operations', 'EMP-5001', 61000],
        [2024, 'Operations', 'EMP-5002', 65000],
        [2024, 'Operations', 'EMP-5003', 58500],
        [2024, 'Operations', 'EMP-5004', 66500],
        [2024, 'Operations', 'EMP-5005', 59000],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Set column widths
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 16
    ws1.column_dimensions['C'].width = 14
    ws1.column_dimensions['D'].width = 16

    # Format salary column as currency
    for row in range(2, len(data) + 2):
        ws1.cell(row=row, column=4).number_format = '$#,##0'

    # Freeze header row
    ws1.freeze_panes = 'A2'

    # ----------------------------------------------------------------
    # Sheet2: Summary (intentionally EMPTY — agent must build this)
    # ----------------------------------------------------------------
    ws2 = wb.create_sheet('Summary')
    # Leave completely empty — agent's task is to build the comparison table here

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
