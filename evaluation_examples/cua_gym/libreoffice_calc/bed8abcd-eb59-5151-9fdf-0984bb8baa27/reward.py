"""
Reward Script: Calculate commission in G2 using nested IF+SUMIF formula
Task ID: calc_fmb_nested_if_sumif_054
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): G2 contains a formula (not empty)
  Component 2 (0.4): Formula uses SUMIF with correct range B2:B201, criteria F2, sum range C2:C201
  Component 3 (0.3): Formula uses IF with threshold >100000, commission rates 0.08 and 0.05
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmb_nested_if_sumif_054'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: verify the 'Commission' sheet exists
    if 'Commission' not in wb.sheetnames:
        print("CRITICAL: 'Commission' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Commission']

    # Component 1: G2 contains a formula (not empty/None) — 0.3 points
    # This FAILS on initial (G2 is None) and PASSES on golden (G2 has formula)
    try:
        g2_value = ws['G2'].value
        if g2_value is not None and isinstance(g2_value, str) and g2_value.startswith('='):
            print(f"PASS: Component 1 — G2 contains a formula: {g2_value[:60]}... (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — G2 should contain a formula, found: {repr(g2_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Formula uses SUMIF with correct arguments — 0.4 points
    # Must reference: B2:B201 (lookup column), F2 (criteria cell), C2:C201 (sum column)
    # This FAILS on initial (G2 is empty) and PASSES on golden (formula has correct SUMIF)
    try:
        g2_value = ws['G2'].value
        if g2_value is not None and isinstance(g2_value, str):
            formula_upper = g2_value.upper().replace(' ', '')
            # Check for SUMIF with expected ranges
            has_sumif = 'SUMIF(' in formula_upper
            has_b_range = 'B2:B201' in formula_upper
            has_f2_criteria = 'F2' in formula_upper
            has_c_range = 'C2:C201' in formula_upper

            if has_sumif and has_b_range and has_f2_criteria and has_c_range:
                print(f"PASS: Component 2 — SUMIF uses correct ranges (B2:B201, F2, C2:C201) (0.4 pts)")
                total_score += 0.4
            else:
                details = []
                if not has_sumif:
                    details.append("missing SUMIF function")
                if not has_b_range:
                    details.append("missing B2:B201 lookup range")
                if not has_f2_criteria:
                    details.append("missing F2 criteria cell")
                if not has_c_range:
                    details.append("missing C2:C201 sum range")
                print(f"FAIL: Component 2 — SUMIF range issue: {', '.join(details)}")
        else:
            print("FAIL: Component 2 — G2 is empty, cannot check SUMIF ranges")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Formula uses IF with correct threshold >100000 and rates 0.08/0.05 — 0.3 points
    # The IF condition must check >100000, with 8% for high earners and 5% for others
    # This FAILS on initial (G2 is empty) and PASSES on golden (formula has correct IF logic)
    try:
        g2_value = ws['G2'].value
        if g2_value is not None and isinstance(g2_value, str):
            formula_upper = g2_value.upper().replace(' ', '')
            # Check IF function
            has_if = 'IF(' in formula_upper
            # Check threshold: >100000
            has_threshold = '>100000' in formula_upper
            # Check commission rates: 0.08 and 0.05
            has_rate_08 = '0.08' in formula_upper or '*8%' in formula_upper
            has_rate_05 = '0.05' in formula_upper or '*5%' in formula_upper

            if has_if and has_threshold and has_rate_08 and has_rate_05:
                print(f"PASS: Component 3 — IF uses correct threshold (>100000) and rates (8%/5%) (0.3 pts)")
                total_score += 0.3
            else:
                details = []
                if not has_if:
                    details.append("missing IF function")
                if not has_threshold:
                    details.append("missing >100000 threshold")
                if not has_rate_08:
                    details.append("missing 0.08 (8%) rate")
                if not has_rate_05:
                    details.append("missing 0.05 (5%) rate")
                print(f"FAIL: Component 3 — IF logic issue: {', '.join(details)}")
        else:
            print("FAIL: Component 3 — G2 is empty, cannot check IF logic")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
