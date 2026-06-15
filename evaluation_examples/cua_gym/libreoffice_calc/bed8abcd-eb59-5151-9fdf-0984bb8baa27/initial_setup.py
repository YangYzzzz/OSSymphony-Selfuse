"""
Initial Setup: Nested IF+SUMIF commission calculation
Task ID: calc_fmb_nested_if_sumif_054
Domain: libreoffice_calc
"""

import os
import openpyxl
from datetime import date, timedelta
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_nested_if_sumif_054'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Commission ---
    ws = wb.active
    ws.title = 'Commission'

    # Row 1: Headers
    ws['A1'] = 'Trans ID'
    ws['B1'] = 'Salesperson'
    ws['C1'] = 'Sale Amount'
    ws['D1'] = 'Date'

    # Salesperson names
    salespersons = [
        'Rachel Adams',
        'Daniel Torres',
        'Priya Nair',
        'James Whitfield',
        'Sofia Hernandez',
        'Kevin Park',
        'Linda Osei',
        'Marcus Reid',
    ]

    # We need Rachel Adams (row 2) and her total across B2:B201 = 127400
    # Design: Rachel gets 10 transactions summing to 127400
    # We'll place Rachel in rows 2, 15, 28, 41, 54, 67, 80, 93, 106, 119
    # with amounts: 18000, 14200, 12500, 11800, 13700, 12600, 11500, 9800, 12300, 11000 = 127400
    rachel_amounts = [18000, 14200, 12500, 11800, 13700, 12600, 11500, 9800, 12300, 11000]
    assert sum(rachel_amounts) == 127400, f"Rachel total mismatch: {sum(rachel_amounts)}"

    rachel_rows = [2, 15, 28, 41, 54, 67, 80, 93, 106, 119]

    # Start date: 2024-01-02
    base_date = date(2024, 1, 2)

    rachel_idx = 0
    trans_id = 1001

    for row_num in range(2, 202):  # rows 2 to 201 inclusive (200 rows)
        trans_row = row_num - 2  # 0-indexed offset
        current_date = base_date + timedelta(days=trans_row)

        if row_num in rachel_rows:
            salesperson = 'Rachel Adams'
            amount = rachel_amounts[rachel_idx]
            rachel_idx += 1
        else:
            # Pick a non-Rachel salesperson deterministically
            idx = (row_num * 7 + 3) % (len(salespersons) - 1) + 1  # indices 1..7
            salesperson = salespersons[idx]
            # Vary amounts realistically: 5000 to 35000
            amount_base = [5200, 7800, 9400, 11200, 6700, 8300, 15400, 21000,
                           13600, 4900, 22500, 17800, 8100, 9900, 6400, 12300,
                           19700, 14200, 7600, 11500, 25000, 8800, 16300, 10100,
                           5500, 18900, 7200, 13800, 9200, 20400]
            amount = amount_base[trans_row % len(amount_base)]

        ws.cell(row=row_num, column=1, value=trans_id)
        ws.cell(row=row_num, column=2, value=salesperson)
        ws.cell(row=row_num, column=3, value=amount)
        ws.cell(row=row_num, column=4, value=current_date.strftime('%Y-%m-%d'))

        trans_id += 1

    # F1 = 'Summary', F2 = 'Rachel Adams', G1 = 'Commission', G2 = empty
    ws['F1'] = 'Summary'
    ws['F2'] = 'Rachel Adams'
    ws['G1'] = 'Commission'
    # G2 is intentionally left empty — this is the target cell

    # Column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Commission')
    print(f'  Rows 2-201: 200 transactions (Rachel Adams total = 127,400)')
    print(f'  G2 is EMPTY (target cell for the task)')

create_initial()
