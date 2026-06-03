"""
Initial Setup: Create a spreadsheet with advertising spend vs sales revenue data.
Task ID: calc_chart_scatter_012
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_scatter_012'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: AdSpend ---
    ws = wb.active
    ws.title = 'AdSpend'

    # Headers
    ws['A1'] = 'Ad Spend ($000)'
    ws['B1'] = 'Sales Revenue ($000)'

    # Data rows (advertising spend and corresponding sales revenue)
    data = [
        (10, 85),
        (15, 102),
        (20, 128),
        (25, 145),
        (30, 168),
        (35, 181),
        (40, 205),
        (45, 219),
    ]

    for r, (ad_spend, sales_rev) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=ad_spend)
        ws.cell(row=r, column=2, value=sales_rev)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 22

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: AdSpend')
    print(f'  Headers: A1=Ad Spend ($000), B1=Sales Revenue ($000)')
    print(f'  Data rows: 2-9 (8 data points)')
    print(f'  No charts (task requires creating one)')


create_initial()
