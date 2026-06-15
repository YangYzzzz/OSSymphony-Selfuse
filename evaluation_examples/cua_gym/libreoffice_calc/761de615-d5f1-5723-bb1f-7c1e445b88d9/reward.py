"""
Reward Script: Freeze first row and first two columns on 'Data' sheet
Task ID: calc_gg3_003
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Freeze panes is set (not None)
  Component 2 (0.4): Freeze panes is exactly 'C2' (row 1 + cols A-B frozen)
  Component 3 (0.2): Freeze panes correct AND data integrity intact
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_003'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify that freeze panes are correctly set on the 'Data' sheet.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: file must exist and be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Data' sheet must exist
    if 'Data' not in wb.sheetnames:
        print(f"CRITICAL: 'Data' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Data']

    # Component 1: Freeze panes is set (not None) — 0.4 points
    # This FAILS on initial (freeze_panes=None), PASSES on golden (freeze_panes='C2')
    try:
        freeze_val = ws.freeze_panes
        if freeze_val is not None:
            print(f"PASS: Component 1 — Freeze panes is set: '{freeze_val}' (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — Freeze panes is None (no freeze applied)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Freeze panes is exactly 'C2' — 0.4 points
    # C2 means: row 1 frozen (header) + columns A and B frozen (Employee ID, Full Name)
    # This FAILS on initial (None != 'C2'), PASSES on golden ('C2' == 'C2')
    try:
        freeze_val = ws.freeze_panes
        if freeze_val is not None and str(freeze_val) == 'C2':
            print(f"PASS: Component 2 — Freeze panes is exactly 'C2' (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 2 — Expected freeze_panes='C2', found: '{freeze_val}'")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Freeze correct AND data integrity intact — 0.2 points
    # Compound check: freeze must be 'C2' AND the data must still be intact
    # (header row present, at least 500 data rows, columns A-B have expected content)
    # This FAILS on initial (freeze is None), PASSES on golden (freeze is C2 + data intact)
    try:
        freeze_val = ws.freeze_panes
        freeze_ok = freeze_val is not None and str(freeze_val) == 'C2'

        if not freeze_ok:
            print(f"FAIL: Component 3 — Freeze panes not 'C2', skipping data integrity check")
        else:
            # Check header row
            header_a = ws.cell(row=1, column=1).value
            header_b = ws.cell(row=1, column=2).value
            headers_ok = (header_a == 'Employee ID' and header_b == 'Full Name')

            # Check data rows exist (should be 500 rows of data, rows 2-501)
            rows_ok = ws.max_row >= 501

            # Check a sample data cell in column A (Employee ID format)
            sample_id = ws.cell(row=2, column=1).value
            id_ok = sample_id is not None and str(sample_id).startswith('EMP-')

            if headers_ok and rows_ok and id_ok:
                print(f"PASS: Component 3 — Freeze correct + data intact: headers='{header_a}','{header_b}', rows={ws.max_row}, sample_id='{sample_id}' (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 3 — Data integrity issues: headers_ok={headers_ok}, rows_ok={rows_ok} (max_row={ws.max_row}), id_ok={id_ok}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
