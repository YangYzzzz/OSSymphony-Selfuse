"""
Initial Setup: Freeze panes task - create employee database spreadsheet
Task ID: calc_gg3_003
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_003'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

# --- Realistic data pools ---
FIRST_NAMES = [
    "Sarah", "Marcus", "Emily", "James", "Priya", "Carlos", "Aisha", "David",
    "Mei", "Robert", "Fatima", "Michael", "Yuki", "Daniel", "Sofia", "Kevin",
    "Olivia", "Raj", "Hannah", "Thomas", "Lena", "Brian", "Zara", "Andrew",
    "Grace", "Nathan", "Chloe", "Victor", "Amara", "Patrick", "Isabella",
    "Jorge", "Nadia", "Ryan", "Leila", "Derek", "Simone", "Trevor", "Aria",
    "Samuel", "Elise", "Oscar", "Mina", "Luke", "Tanya", "Felix", "Diana",
    "Hugo", "Clara", "Sean"
]

LAST_NAMES = [
    "Chen", "Johnson", "Patel", "Williams", "Kim", "Rodriguez", "O'Brien",
    "Singh", "Taylor", "Nakamura", "Hassan", "Brown", "Kowalski", "Davis",
    "Fernandez", "Wilson", "Nguyen", "Martinez", "Anderson", "Thompson",
    "Garcia", "Lee", "Robinson", "Clark", "Lewis", "Walker", "Hall",
    "Young", "Wright", "Lopez", "Hill", "Scott", "Green", "Adams",
    "Baker", "Nelson", "Carter", "Mitchell", "Perez", "Roberts",
    "Turner", "Phillips", "Campbell", "Parker", "Evans", "Edwards",
    "Collins", "Stewart", "Sanchez", "Morris"
]

DEPARTMENTS = [
    "Engineering", "Marketing", "Sales", "Human Resources", "Finance",
    "Operations", "Legal", "Product", "Customer Success", "Research",
    "IT Support", "Data Science", "Design", "Business Development", "Quality Assurance"
]

TITLES = [
    "Analyst", "Senior Analyst", "Associate", "Senior Associate",
    "Specialist", "Senior Specialist", "Manager", "Senior Manager",
    "Director", "Vice President", "Coordinator", "Lead", "Engineer",
    "Senior Engineer", "Principal Engineer", "Consultant", "Architect"
]

CITIES = [
    ("New York", "NY", "US"), ("San Francisco", "CA", "US"),
    ("Chicago", "IL", "US"), ("Austin", "TX", "US"),
    ("Seattle", "WA", "US"), ("Boston", "MA", "US"),
    ("Denver", "CO", "US"), ("Atlanta", "GA", "US"),
    ("Portland", "OR", "US"), ("Miami", "FL", "US"),
    ("London", "", "UK"), ("Toronto", "ON", "CA"),
    ("Berlin", "", "DE"), ("Singapore", "", "SG"),
    ("Sydney", "NSW", "AU"), ("Dublin", "", "IE"),
    ("Amsterdam", "", "NL"), ("Tokyo", "", "JP"),
    ("Mumbai", "MH", "IN"), ("Sao Paulo", "SP", "BR"),
]

OFFICES = ["HQ-Floor 1", "HQ-Floor 2", "HQ-Floor 3", "East Wing", "West Wing",
           "Building A", "Building B", "Remote", "Satellite-North", "Satellite-South"]

STATUSES = ["Active", "Active", "Active", "Active", "Active", "Active",
            "On Leave", "Probation", "Active", "Active"]

random.seed(42)

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # --- Headers in row 1 ---
    headers = [
        "Employee ID", "Full Name", "Department", "Title", "Salary",
        "Start Date", "Email", "Phone", "Office", "City",
        "State", "Country", "Manager", "Rating", "Bonus", "Status"
    ]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- 500 rows of employee data (rows 2-501) ---
    managers = []
    for i in range(1, 501):
        row = i + 1
        emp_id = f"EMP-{10000 + i}"
        first = random.choice(FIRST_NAMES)
        last = random.choice(LAST_NAMES)
        full_name = f"{first} {last}"
        dept = random.choice(DEPARTMENTS)
        title = random.choice(TITLES)
        salary = round(random.uniform(45000, 185000), 2)

        # Start date between 2015-01-01 and 2025-12-31
        year = random.randint(2015, 2025)
        month = random.randint(1, 12)
        day = random.randint(1, 28)
        start_date = f"{year}-{month:02d}-{day:02d}"

        last_clean = last.lower().replace("'", "")
        email = f"{first.lower()}.{last_clean}@globalcorp.com"
        phone_area = random.randint(200, 999)
        phone_mid = random.randint(100, 999)
        phone_end = random.randint(1000, 9999)
        phone = f"({phone_area}) {phone_mid}-{phone_end}"

        office = random.choice(OFFICES)
        city_info = random.choice(CITIES)
        city, state, country = city_info

        if managers and random.random() > 0.15:
            manager = random.choice(managers[:min(len(managers), 50)])
        else:
            manager = "N/A"
        if "Manager" in title or "Director" in title or "VP" in title:
            managers.append(full_name)

        rating = round(random.uniform(2.0, 5.0), 1)
        bonus = round(salary * random.uniform(0.02, 0.20), 2)
        status = random.choice(STATUSES)

        row_data = [
            emp_id, full_name, dept, title, salary, start_date,
            email, phone, office, city, state, country, manager,
            rating, bonus, status
        ]
        for c, val in enumerate(row_data, 1):
            ws.cell(row=row, column=c, value=val)

    # --- Column widths ---
    col_widths = {
        'A': 14, 'B': 22, 'C': 18, 'D': 22, 'E': 14,
        'F': 14, 'G': 32, 'H': 18, 'I': 16, 'J': 16,
        'K': 10, 'L': 10, 'M': 22, 'N': 10, 'O': 14, 'P': 12
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Row 1 height
    ws.row_dimensions[1].height = 22

    # NO freeze panes - this is the task the agent must perform
    # ws.freeze_panes is left as None

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
