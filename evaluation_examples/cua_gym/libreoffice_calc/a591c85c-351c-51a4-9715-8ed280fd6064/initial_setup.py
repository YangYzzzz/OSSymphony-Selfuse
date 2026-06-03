"""
Initial Setup: Name cell B2 as 'ExchangeRate' and write EUR conversion formula
Task ID: calc_nrv_037
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_037'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # --- Column A & B: Exchange rate info ---
    ws["A1"] = "Item"
    ws["B1"] = "EUR/USD Rate"
    ws["B2"] = 1.08

    # Item names in A2:A10
    items = [
        "Wireless Bluetooth Headphones",
        "USB-C Charging Cable",
        "Laptop Stand Adjustable",
        "Mechanical Keyboard RGB",
        "Portable External SSD 1TB",
        "Webcam HD 1080p",
        "Mouse Pad XL Gaming",
        "Monitor Arm Single",
        "Desk Organizer Set",
    ]
    for i, item in enumerate(items, 2):
        ws.cell(row=i, column=1, value=item)

    # --- Column D & E: Price conversion area ---
    ws["D1"] = "Price USD"
    ws["E1"] = "Price EUR"

    # USD prices in D2:D10
    usd_prices = [49.99, 12.95, 89.50, 134.99, 109.00, 64.75, 19.99, 75.00, 34.50]
    for i, price in enumerate(usd_prices, 2):
        ws.cell(row=i, column=4, value=price)

    # E2 is intentionally left EMPTY — the agent must fill it
    # No named ranges defined — the agent must create 'ExchangeRate'

    # Style headers for a professional look
    header_font = Font(bold=True)
    for cell_ref in ["A1", "B1", "D1", "E1"]:
        ws[cell_ref].font = header_font

    # Set reasonable column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
