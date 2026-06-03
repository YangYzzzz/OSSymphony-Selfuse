"""
Initial Setup: Create a workbook with Index and Inventory sheets for dynamic INDIRECT formula task.
Task ID: calc_mcp_055
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_055'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Index ---
    ws_index = wb.active
    ws_index.title = 'Index'

    # Headers
    headers = ['Sheet Name', 'Result', 'Row Number']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws_index.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data row: A2=Inventory, B2=empty (agent must fill this), C2=15
    ws_index.cell(row=2, column=1, value='Inventory')
    # B2 intentionally left empty - this is what the task asks the agent to create
    ws_index.cell(row=2, column=3, value=15)

    # Additional reference rows to make it more realistic
    ws_index.cell(row=3, column=1, value='Inventory')
    ws_index.cell(row=3, column=3, value=30)
    # B3 also empty

    ws_index.cell(row=4, column=1, value='Inventory')
    ws_index.cell(row=4, column=3, value=7)
    # B4 also empty

    # Column widths
    ws_index.column_dimensions['A'].width = 18
    ws_index.column_dimensions['B'].width = 16
    ws_index.column_dimensions['C'].width = 14

    # --- Sheet 2: Inventory ---
    ws_inv = wb.create_sheet('Inventory')

    # Headers
    inv_headers = ['Item', 'Value', 'Category', 'In Stock']
    for col, h in enumerate(inv_headers, 1):
        cell = ws_inv.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Realistic inventory data for rows 2-50 (B1:B50 populated)
    import random
    random.seed(42)

    items = [
        ('Wireless Mouse', 29.99, 'Electronics', True),
        ('USB-C Hub', 45.50, 'Electronics', True),
        ('Desk Lamp', 34.75, 'Office', True),
        ('Ergonomic Keyboard', 89.99, 'Electronics', False),
        ('Monitor Stand', 52.00, 'Office', True),
        ('Webcam HD', 67.25, 'Electronics', True),
        ('Cable Organizer', 12.99, 'Office', True),
        ('Noise-Cancelling Headset', 149.99, 'Electronics', False),
        ('Whiteboard Markers (12pk)', 18.50, 'Supplies', True),
        ('Standing Desk Mat', 44.00, 'Office', True),
        ('Portable Charger', 35.99, 'Electronics', True),
        ('Laptop Sleeve 15"', 24.75, 'Accessories', True),
        ('Mechanical Pencil Set', 8.99, 'Supplies', True),
        ('Bluetooth Speaker', 59.99, 'Electronics', False),
        ('Desk Organizer Tray', 22.50, 'Office', True),
        ('HDMI Cable 6ft', 14.99, 'Electronics', True),
        ('Sticky Notes (Assorted)', 6.75, 'Supplies', True),
        ('Laptop Cooling Pad', 31.00, 'Electronics', True),
        ('Paper Shredder', 78.50, 'Office', False),
        ('External SSD 500GB', 69.99, 'Electronics', True),
        ('Binder Clips (50pk)', 5.25, 'Supplies', True),
        ('Adjustable Monitor Arm', 119.99, 'Office', True),
        ('Wireless Charger', 27.50, 'Electronics', True),
        ('Presentation Remote', 32.99, 'Electronics', True),
        ('Filing Cabinet Label', 3.99, 'Supplies', True),
        ('USB Flash Drive 64GB', 11.50, 'Electronics', True),
        ('Desk Calendar 2025', 9.99, 'Supplies', False),
        ('Ethernet Cable 10ft', 8.75, 'Electronics', True),
        ('Paper Clips (200pk)', 4.50, 'Supplies', True),
        ('Surge Protector', 24.99, 'Electronics', True),
        ('Tape Dispenser', 7.25, 'Supplies', True),
        ('Webcam Cover', 3.50, 'Accessories', True),
        ('Mouse Pad XL', 19.99, 'Accessories', True),
        ('Screen Cleaning Kit', 12.50, 'Supplies', True),
        ('Cable Ties (100pk)', 6.99, 'Supplies', True),
        ('Portable Scanner', 95.00, 'Electronics', False),
        ('Desk Phone Stand', 15.75, 'Accessories', True),
        ('Label Maker', 42.50, 'Office', True),
        ('Anti-Glare Screen', 28.99, 'Accessories', True),
        ('Document Holder', 17.25, 'Office', True),
        ('Pen Holder', 10.50, 'Office', True),
        ('Footrest', 38.99, 'Office', True),
        ('Wrist Rest', 14.25, 'Accessories', True),
        ('Privacy Screen Filter', 33.50, 'Accessories', False),
        ('Desk Drawer Organizer', 21.75, 'Office', True),
        ('Compressed Air Can', 8.99, 'Supplies', True),
        ('USB Microphone', 55.00, 'Electronics', True),
        ('Desk Shelf Riser', 29.50, 'Office', True),
        ('Power Strip 8-Outlet', 19.99, 'Electronics', True),
    ]

    for r, (item, value, category, in_stock) in enumerate(items, 2):
        ws_inv.cell(row=r, column=1, value=item)
        ws_inv.cell(row=r, column=2, value=value)
        ws_inv.cell(row=r, column=3, value=category)
        ws_inv.cell(row=r, column=4, value='Yes' if in_stock else 'No')

    # Column widths for Inventory
    ws_inv.column_dimensions['A'].width = 28
    ws_inv.column_dimensions['B'].width = 12
    ws_inv.column_dimensions['C'].width = 16
    ws_inv.column_dimensions['D'].width = 12

    # Freeze header row on Inventory
    ws_inv.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Note: Row 15 in Inventory (B15) has value 22.50 (Desk Organizer Tray)
    # This is what INDIRECT(A2&".B"&C2) should resolve to when A2=Inventory, C2=15

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
