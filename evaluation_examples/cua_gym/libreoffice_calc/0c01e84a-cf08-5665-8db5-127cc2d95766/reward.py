"""
Reward Script: Dynamic INDIRECT formula in Index!B2
Task ID: calc_mcp_055
Domain: libreoffice_calc
Scoring:
  Component 1 (0.40): B2 contains a formula (not empty/literal)
  Component 2 (0.35): Formula uses INDIRECT to build dynamic reference
  Component 3 (0.25): Formula references A2 (sheet name) and C2 (row number) correctly
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_055'


def persist_app_state(domain):
    """Save any unsaved LibreOffice changes before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print(f"PERSIST: ctrl+s sent for {domain}")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Index' sheet must exist
    if 'Index' not in wb.sheetnames:
        print("FAIL: 'Index' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Index']
    b2_value = ws['B2'].value

    # Component 1: B2 contains a formula (0.40 points)
    # In the initial env, B2 is empty (None). A formula is a string starting with '='.
    try:
        if b2_value is not None and isinstance(b2_value, str) and b2_value.startswith('='):
            print(f"PASS: Component 1 -- B2 contains a formula: {b2_value} (0.40 pts)")
            total_score += 0.40
        else:
            print(f"FAIL: Component 1 -- B2 does not contain a formula, found: {repr(b2_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: The formula uses INDIRECT function (0.35 points)
    # The task specifically asks for a dynamic formula using INDIRECT.
    try:
        if b2_value is not None and isinstance(b2_value, str):
            formula_upper = b2_value.upper().replace(" ", "")
            if 'INDIRECT(' in formula_upper:
                print(f"PASS: Component 2 -- Formula uses INDIRECT function (0.35 pts)")
                total_score += 0.35
            else:
                print(f"FAIL: Component 2 -- Formula does not use INDIRECT, found: {b2_value}")
        else:
            print(f"FAIL: Component 2 -- B2 is not a formula string")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Formula references A2 (sheet name) and C2 (row number) to build
    # the dynamic cross-sheet reference (0.25 points)
    # Valid formulas include patterns like:
    #   =INDIRECT(A2&".B"&C2)
    #   =INDIRECT(A2&"!B"&C2)
    #   =INDIRECT(A2&".B"&C2)  (LibreOffice dot notation)
    # The key requirement: references A2 for sheet name and C2 for row number
    try:
        if b2_value is not None and isinstance(b2_value, str):
            formula_upper = b2_value.upper().replace(" ", "")
            has_a2_ref = bool(re.search(r'A2', formula_upper))
            has_c2_ref = bool(re.search(r'C2', formula_upper))
            # Check that column B is referenced in the constructed string
            has_col_b = bool(re.search(r'["\']\.?B["\']|["\']!?B["\']|"\.B"|"!B"', b2_value, re.IGNORECASE))
            if has_a2_ref and has_c2_ref and has_col_b:
                print(f"PASS: Component 3 -- Formula references A2, column B, and C2 correctly (0.25 pts)")
                total_score += 0.25
            else:
                details = f"A2 ref: {has_a2_ref}, C2 ref: {has_c2_ref}, col B ref: {has_col_b}"
                print(f"FAIL: Component 3 -- Incomplete dynamic reference. {details}. Formula: {b2_value}")
        else:
            print(f"FAIL: Component 3 -- B2 is not a formula string")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
