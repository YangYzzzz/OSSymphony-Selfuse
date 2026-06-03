"""
Initial Setup: Create a workbook with 20 department sheets (Dept_01 to Dept_20).
Task ID: calc_ps_092
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_092'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

# Realistic department data per sheet
DEPT_NAMES = {
    'Dept_01': 'Human Resources',
    'Dept_02': 'Finance',
    'Dept_03': 'Engineering',
    'Dept_04': 'Marketing',
    'Dept_05': 'Sales',
    'Dept_06': 'Customer Support',
    'Dept_07': 'Legal',
    'Dept_08': 'Operations',
    'Dept_09': 'Research & Development',
    'Dept_10': 'Quality Assurance',
    'Dept_11': 'Procurement',
    'Dept_12': 'Logistics',
    'Dept_13': 'IT Infrastructure',
    'Dept_14': 'Data Analytics',
    'Dept_15': 'Product Management',
    'Dept_16': 'Design',
    'Dept_17': 'Training',
    'Dept_18': 'Compliance',
    'Dept_19': 'Facilities',
    'Dept_20': 'Executive Office',
}

EMPLOYEE_POOLS = [
    ['Sarah Chen', 'Senior Specialist', 92000, '2021-03-15'],
    ['Marcus Johnson', 'Team Lead', 105000, '2019-07-22'],
    ['Emily Rodriguez', 'Analyst', 68000, '2023-01-10'],
    ['David Kim', 'Manager', 115000, '2018-11-05'],
    ['Aisha Patel', 'Coordinator', 58000, '2024-02-18'],
    ['James O\'Brien', 'Director', 135000, '2017-06-30'],
    ['Lisa Wang', 'Associate', 62000, '2023-09-01'],
    ['Robert Martinez', 'Consultant', 88000, '2020-04-12'],
    ['Nadia Okafor', 'Engineer', 97000, '2022-08-20'],
    ['Thomas Fischer', 'Supervisor', 82000, '2021-12-03'],
    ['Priya Sharma', 'Specialist', 73000, '2022-05-14'],
    ['Carlos Mendez', 'Intern', 42000, '2025-01-06'],
]

HEADERS = ['Employee Name', 'Role', 'Annual Salary', 'Start Date']

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    wb = openpyxl.Workbook()

    header_font = Font(name='Arial', size=11, bold=True)
    header_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for idx, (sheet_key, dept_full_name) in enumerate(DEPT_NAMES.items()):
        if idx == 0:
            ws = wb.active
            ws.title = sheet_key
        else:
            ws = wb.create_sheet(sheet_key)

        # Write department title in row 1
        ws.cell(row=1, column=1, value=f'{dept_full_name} Department')
        ws.cell(row=1, column=1).font = Font(name='Arial', size=14, bold=True)
        ws.merge_cells('A1:D1')
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        # Headers in row 3
        for col, h in enumerate(HEADERS, 1):
            c = ws.cell(row=3, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = header_align
            c.border = thin_border

        # Rotate employee data based on dept index for variety
        num_employees = 8 + (idx % 5)  # 8-12 employees per dept
        for r in range(num_employees):
            emp = EMPLOYEE_POOLS[(r + idx * 3) % len(EMPLOYEE_POOLS)]
            for col, val in enumerate(emp, 1):
                cell = ws.cell(row=4 + r, column=col, value=val)
                cell.border = thin_border

        # Column widths
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 14

    # Ensure Dept_01 is the active sheet
    wb.active = wb.worksheets[0]

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
