"""
Reward Script: Set column A width to 5cm and column B width to 2cm
Task ID: calc_gfl_076
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.4): Column A width ~5cm (18.5-18.7 char units)
  - Component 2 (0.4): Column B width ~2cm (7.3-7.5 char units)
  - Component 3 (0.2): Data integrity — content unchanged, customWidth set
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_076'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Column width conversion: LibreOffice Calc uses cm in the Column Width dialog.
    openpyxl stores width in character units. The approximate conversion is:
        width_chars ≈ cm * 3.714
    So 5cm ≈ 18.57 chars, 2cm ≈ 7.43 chars.
    We allow a tolerance of ±0.5 character units to account for rounding.
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Weekly' sheet must exist
    if 'Weekly' not in wb.sheetnames:
        print("CRITICAL: 'Weekly' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Weekly']

    # Expected column widths (in openpyxl character units)
    # 5cm ≈ 18.57 chars, 2cm ≈ 7.43 chars
    EXPECTED_A_WIDTH = 18.57
    EXPECTED_B_WIDTH = 7.43
    TOLERANCE = 0.5  # allow ±0.5 char units

    # Component 1: Column A width is approximately 5cm (0.4 points)
    try:
        col_a = ws.column_dimensions.get('A')
        if col_a is not None and col_a.width is not None:
            a_width = col_a.width
            if abs(a_width - EXPECTED_A_WIDTH) <= TOLERANCE:
                print(f"PASS: Component 1 — Column A width={a_width:.2f}, expected ~{EXPECTED_A_WIDTH} (0.4 pts)")
                total_score += 0.4
            else:
                print(f"FAIL: Component 1 — Column A width={a_width:.2f}, expected ~{EXPECTED_A_WIDTH} (tolerance ±{TOLERANCE})")
        else:
            print(f"FAIL: Component 1 — Column A has no custom width set (still default)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Column B width is approximately 2cm (0.4 points)
    try:
        col_b = ws.column_dimensions.get('B')
        if col_b is not None and col_b.width is not None:
            b_width = col_b.width
            if abs(b_width - EXPECTED_B_WIDTH) <= TOLERANCE:
                print(f"PASS: Component 2 — Column B width={b_width:.2f}, expected ~{EXPECTED_B_WIDTH} (0.4 pts)")
                total_score += 0.4
            else:
                print(f"FAIL: Component 2 — Column B width={b_width:.2f}, expected ~{EXPECTED_B_WIDTH} (tolerance ±{TOLERANCE})")
        else:
            print(f"FAIL: Component 2 — Column B has no custom width set (still default)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Data integrity and customWidth flags (0.2 points)
    # Verify that both columns have customWidth=True AND data is unchanged
    try:
        sub_score = 0.0

        # Check customWidth on column A
        col_a = ws.column_dimensions.get('A')
        col_b = ws.column_dimensions.get('B')
        a_custom = col_a is not None and col_a.customWidth
        b_custom = col_b is not None and col_b.customWidth

        if a_custom and b_custom:
            sub_score += 0.1
            print(f"PASS: Component 3a — Both columns have customWidth=True")
        else:
            print(f"FAIL: Component 3a — customWidth: A={a_custom}, B={b_custom}")

        # Verify data integrity: check key cells haven't changed
        expected_a1 = "Activity Name"
        expected_b1 = "Code"
        expected_a2 = "Advanced Cardiovascular Fitness Training"
        expected_b2 = "ACF"

        a1_val = ws['A1'].value
        b1_val = ws['B1'].value
        a2_val = ws['A2'].value
        b2_val = ws['B2'].value

        data_ok = (
            str(a1_val).strip() == expected_a1
            and str(b1_val).strip() == expected_b1
            and str(a2_val).strip() == expected_a2
            and str(b2_val).strip() == expected_b2
        )

        if data_ok:
            sub_score += 0.1
            print(f"PASS: Component 3b — Data integrity verified (headers and sample data match)")
        else:
            print(f"FAIL: Component 3b — Data changed: A1={a1_val}, B1={b1_val}, A2={a2_val}, B2={b2_val}")

        # Only award component 3 if at least one column width was changed (components 1 or 2 passed)
        # This ensures component 3 doesn't award points on the initial env where no width changes were made
        if total_score > 0 and sub_score > 0:
            total_score += sub_score
            print(f"PASS: Component 3 — total sub_score={sub_score:.1f} (0.2 pts)")
        elif total_score == 0:
            print(f"FAIL: Component 3 — gated: no width changes detected, skipping data integrity scoring")
        elif sub_score > 0:
            total_score += sub_score
            print(f"PARTIAL: Component 3 — sub_score={sub_score:.1f}")
        else:
            print(f"FAIL: Component 3 — no sub-checks passed")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
