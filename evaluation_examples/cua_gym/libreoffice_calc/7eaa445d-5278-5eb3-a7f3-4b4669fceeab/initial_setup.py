"""
Initial Setup: Weekly schedule spreadsheet with default column widths
Task ID: calc_gfl_076
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_076'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Weekly ---
    ws = wb.active
    ws.title = 'Weekly'

    # Headers (8 columns)
    headers = [
        'Activity Name', 'Code', 'Day', 'Start Time',
        'End Time', 'Duration (hrs)', 'Location', 'Instructor'
    ]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Data rows (24 rows of realistic schedule data)
    # Column A has long activity names that will be truncated at default width
    data = [
        ['Advanced Cardiovascular Fitness Training', 'ACF', 'Monday', '06:00', '07:30', 1.5, 'Gymnasium Hall A', 'Sarah Chen'],
        ['Beginner Yoga and Mindfulness Session', 'BYM', 'Monday', '08:00', '09:00', 1.0, 'Studio 3 West Wing', 'Priya Sharma'],
        ['Intermediate Swimming Technique Workshop', 'IST', 'Monday', '09:30', '11:00', 1.5, 'Aquatic Center Pool B', 'Marcus Johnson'],
        ['Strength and Conditioning Bootcamp', 'SCB', 'Monday', '11:30', '12:30', 1.0, 'Outdoor Training Field', 'Jake Morrison'],
        ['Pilates Core Strengthening Fundamentals', 'PCF', 'Tuesday', '06:30', '07:30', 1.0, 'Studio 1 Main Building', 'Elena Rodriguez'],
        ['High-Intensity Interval Training Circuit', 'HIT', 'Tuesday', '08:00', '09:00', 1.0, 'Gymnasium Hall B', 'David Park'],
        ['Senior Wellness and Flexibility Program', 'SWF', 'Tuesday', '09:30', '10:30', 1.0, 'Community Room East', 'Linda Nakamura'],
        ['CrossFit Endurance and Power Session', 'CEP', 'Tuesday', '11:00', '12:30', 1.5, 'CrossFit Zone Alpha', 'Ryan OConnor'],
        ['Meditation and Breathing Techniques Class', 'MBT', 'Wednesday', '06:00', '07:00', 1.0, 'Zen Garden Pavilion', 'Aiko Tanaka'],
        ['Functional Movement Assessment Clinic', 'FMA', 'Wednesday', '07:30', '09:00', 1.5, 'Sports Medicine Lab', 'Dr. Nina Patel'],
        ['Indoor Rock Climbing Fundamentals Course', 'IRC', 'Wednesday', '09:30', '11:00', 1.5, 'Climbing Wall Center', 'Tom Bergstrom'],
        ['Aqua Aerobics for Joint Rehabilitation', 'AAJ', 'Wednesday', '11:30', '12:30', 1.0, 'Therapy Pool Section C', 'Hannah Williams'],
        ['Martial Arts Self-Defense Fundamentals', 'MAS', 'Thursday', '06:00', '07:30', 1.5, 'Dojo Training Room', 'Kenji Watanabe'],
        ['Dance Cardio and Rhythm Fitness Class', 'DCR', 'Thursday', '08:00', '09:00', 1.0, 'Dance Studio North', 'Maria Santos'],
        ['Outdoor Trail Running Preparation Group', 'OTR', 'Thursday', '09:30', '11:00', 1.5, 'Track and Field Area', 'Chris Larsen'],
        ['Rehabilitation Exercise Therapy Session', 'RET', 'Thursday', '11:30', '12:30', 1.0, 'Physiotherapy Suite 2', 'Dr. James Liu'],
        ['Olympic Weightlifting Technique Seminar', 'OWT', 'Friday', '06:30', '08:00', 1.5, 'Weightlifting Platform A', 'Viktor Petrov'],
        ['Spin Cycling High Endurance Challenge', 'SCH', 'Friday', '08:30', '09:30', 1.0, 'Cycling Studio Level 2', 'Sophie Dubois'],
        ['Flexibility and Mobility Recovery Workshop', 'FMR', 'Friday', '10:00', '11:00', 1.0, 'Recovery Lounge South', 'Amara Okafor'],
        ['Team Sports Coordination Drills Practice', 'TSC', 'Friday', '11:30', '13:00', 1.5, 'Multi-Sport Court B', 'Alex Hernandez'],
        ['Weekend Hiking Preparation and Planning', 'WHP', 'Saturday', '07:00', '08:30', 1.5, 'Briefing Room 4', 'Rachel Kim'],
        ['Family Fitness Fun Activity Morning', 'FFF', 'Saturday', '09:00', '10:30', 1.5, 'Gymnasium Hall A', 'Ben Okafor'],
        ['Competitive Badminton Skills Training', 'CBS', 'Saturday', '11:00', '12:30', 1.5, 'Badminton Courts 1-4', 'Mei Lin Wong'],
        ['Restorative Gentle Yoga Evening Session', 'RGY', 'Saturday', '16:00', '17:00', 1.0, 'Studio 3 West Wing', 'Priya Sharma'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # DO NOT set any custom column widths - leave at defaults
    # The task requires the agent to change column A to 5cm and column B to 2cm

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
