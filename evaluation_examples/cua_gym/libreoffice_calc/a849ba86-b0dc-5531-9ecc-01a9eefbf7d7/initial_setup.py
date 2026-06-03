"""
Initial Setup: employees.xlsx with missing salary values open in LibreOffice Calc
Task ID: osworld_multi_apps_calc_vscode_008
Domain: libreoffice_calc + vscode (multi-app)

Creates:
  - /home/user/Desktop/employees.xlsx  (15-row employees spreadsheet with some missing Salary cells)

Opens:
  - LibreOffice Calc showing employees.xlsx
  - VSCode showing Desktop folder (ready for the agent to create a Python script)
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user/Desktop'
TASK_ID = 'osworld_multi_apps_calc_vscode_008'
OUTPUT = f'{WORKDIR}/employees.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    os.makedirs(WORKDIR, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Employees'

    # Headers
    headers = ['Name', 'Department', 'Salary', 'StartDate']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Employee data — some Salary cells intentionally missing (None)
    # Departments: Engineering, Marketing, Sales, HR, Finance
    employees = [
        ('Sarah Chen',          'Engineering', 95000,  '2021-03-15'),
        ('Marcus Johnson',      'Marketing',   72000,  '2020-07-01'),
        ('Emily Rodriguez',     'Sales',       58000,  '2022-01-10'),
        ('David Kim',           'Engineering', None,   '2023-06-20'),  # missing salary
        ('Lisa Thompson',       'HR',          62000,  '2019-11-05'),
        ('James Wilson',        'Finance',     88000,  '2018-04-12'),
        ('Rachel Martinez',     'Marketing',   None,   '2022-09-01'),  # missing salary
        ('Kevin Park',          'Engineering', 102000, '2020-02-28'),
        ('Amanda Foster',       'Sales',       54000,  '2023-03-14'),
        ('Brian Lee',           'Finance',     91000,  '2017-08-22'),
        ('Jennifer Adams',      'HR',          59000,  '2021-05-30'),
        ('Michael Brown',       'Sales',       None,   '2022-11-11'),  # missing salary
        ('Stephanie Turner',    'Engineering', 98000,  '2019-09-09'),
        ('Carlos Rivera',       'Marketing',   76000,  '2021-01-17'),
        ('Nicole Davis',        'Finance',     85000,  '2020-06-03'),
        ('Andrew Chen',         'HR',          64000,  '2018-12-21'),
        ('Melissa Garcia',      'Sales',       61000,  '2023-08-08'),
        ('Robert Taylor',       'Engineering', None,   '2024-01-15'),  # missing salary
        ('Patricia Moore',      'Marketing',   69000,  '2022-04-25'),
        ('Christopher White',   'Finance',     93000,  '2016-10-31'),
    ]

    for r, (name, dept, salary, start_date) in enumerate(employees, 2):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=dept)
        if salary is not None:
            ws.cell(row=r, column=3, value=salary)
        # Leave salary cell empty if None (missing)
        ws.cell(row=r, column=4, value=start_date)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)

    # Also open VSCode with Desktop folder so agent can write Python script there
    launch_gui(f'code "{WORKDIR}"', delay_sec=2.0)

    print('GUI_READY: launched LibreOffice Calc and VSCode with DISPLAY=:0')


create_initial()
