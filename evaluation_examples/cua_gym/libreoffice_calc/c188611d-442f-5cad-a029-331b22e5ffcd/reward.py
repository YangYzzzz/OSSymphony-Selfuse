"""
Reward Script: Classify contracts by urgency using nested IF formulas
Task ID: calc_fma_nested_if_date_050
Domain: libreoffice_calc
Scoring:
  Precondition gate: Sheet 'ContractStatus' exists, headers intact, columns A/B valid
  Component 1 (0.6): All 12 cells C2:C13 contain non-empty IF formulas
  Component 2 (0.4): All 12 formulas have correct structure — uses TODAY(),
                     correct thresholds (<=30, <=90, <=365), and all 4 category
                     labels: "Critical", "Warning", "OK", "Long Term"
"""

import os
import datetime
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fma_nested_if_date_050'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.

    The task requires:
    - C2:C13 to each contain a nested IF formula that classifies contract urgency
    - Formula must compare the row's expiry date (column B) against TODAY(),
      and return one of: 'Critical' (<=30 days), 'Warning' (<=90 days),
      'OK' (<=365 days), 'Long Term' (>365 days)
    - No other cells should be modified

    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: sheet must exist
    if 'ContractStatus' not in wb.sheetnames:
        print("CRITICAL: Sheet 'ContractStatus' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['ContractStatus']

    # Precondition gate: verify basic structure (not scored — present in both initial and golden)
    expected_headers = {1: 'ContractID', 2: 'ExpiryDate', 3: 'Urgency'}
    for col, expected in expected_headers.items():
        actual = ws.cell(row=1, column=col).value
        if actual != expected:
            print(f"CRITICAL: Header check — col {col} expected '{expected}', got '{repr(actual)}'")
            print("REWARD: 0.0")
            return 0.0

    # Precondition gate: columns A and B must be present and valid
    for row in range(2, 14):
        val_a = ws.cell(row=row, column=1).value
        val_b = ws.cell(row=row, column=2).value
        if val_a is None or not str(val_a).startswith('CTR-'):
            print(f"CRITICAL: A{row} should be a contract ID starting with 'CTR-', got {repr(val_a)}")
            print("REWARD: 0.0")
            return 0.0
        if not isinstance(val_b, (datetime.datetime, datetime.date)):
            print(f"CRITICAL: B{row} should be a date, got {repr(val_b)}")
            print("REWARD: 0.0")
            return 0.0
    print("Precondition gate: sheet structure OK (headers, columns A/B intact)")

    # -----------------------------------------------------------------------
    # Component 1: All 12 formula cells C2:C13 are non-empty and contain
    #              IF formulas (0.6 points)
    # This FAILS on the initial file (C2:C13 are all empty) and PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        cells_with_formula = 0
        cells_with_if = 0
        missing_cells = []

        for row in range(2, 14):  # rows 2 through 13
            cell = ws.cell(row=row, column=3)
            val = cell.value
            if val is not None and str(val).strip() != '':
                cells_with_formula += 1
                # Check if the cell value is a formula containing IF
                if isinstance(val, str) and val.startswith('=') and 'IF(' in val.upper():
                    cells_with_if += 1
                else:
                    print(f"  NOTE: C{row} has value but not an IF formula: {repr(val)}")
            else:
                missing_cells.append(f"C{row}")

        if missing_cells:
            print(f"FAIL: Component 1 — {len(missing_cells)} cells still empty: {missing_cells}")
        elif cells_with_if == 12:
            print(f"PASS: Component 1 — All 12 cells C2:C13 contain IF formulas (0.6 pts)")
            total_score += 0.6
        elif cells_with_formula == 12:
            # All cells filled but not all are IF formulas — partial credit
            fraction = cells_with_if / 12
            partial = round(0.6 * fraction, 2)
            print(f"PARTIAL: Component 1 — {cells_with_formula}/12 cells filled, "
                  f"{cells_with_if}/12 are IF formulas ({partial} pts)")
            total_score += partial
        else:
            # Some cells filled — partial credit proportional to filled cells
            fraction = cells_with_formula / 12
            partial = round(0.6 * fraction, 2)
            print(f"PARTIAL: Component 1 — Only {cells_with_formula}/12 cells filled "
                  f"({cells_with_if} with IF formulas) — {partial} pts")
            total_score += partial
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Formula structure correctness — uses TODAY(), correct
    #              thresholds (<=30, <=90, <=365), correct category labels
    #              ("Critical", "Warning", "OK", "Long Term") (0.4 points)
    # This FAILS on the initial file (no formulas in C column at all) and
    # PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        count_today = 0
        count_thresholds = 0
        count_labels = 0

        for row in range(2, 14):
            cell = ws.cell(row=row, column=3)
            val = cell.value
            if not isinstance(val, str) or not val.startswith('='):
                # No formula — all sub-checks fail for this row
                continue

            formula_upper = val.upper().replace(' ', '')
            # Check TODAY() is used
            if 'TODAY()' in formula_upper:
                count_today += 1
            # Check all three numeric thresholds appear
            if '<=30' in formula_upper and '<=90' in formula_upper and '<=365' in formula_upper:
                count_thresholds += 1
            # Check all four category labels — do NOT strip spaces (Long Term has a space)
            if ('"Critical"' in val and '"Warning"' in val
                    and '"OK"' in val and '"Long Term"' in val):
                count_labels += 1

        # All 12 cells should pass each check
        perfect_today = count_today == 12
        perfect_thresholds = count_thresholds == 12
        perfect_labels = count_labels == 12

        if perfect_today and perfect_thresholds and perfect_labels:
            print(f"PASS: Component 2 — All 12 formulas have correct structure: "
                  f"TODAY(), thresholds (30/90/365), all 4 category labels (0.4 pts)")
            total_score += 0.4
        else:
            issues = []
            if not perfect_today:
                issues.append(f"TODAY() found in only {count_today}/12 formulas")
            if not perfect_thresholds:
                issues.append(f"correct thresholds (<=30, <=90, <=365) found in only "
                              f"{count_thresholds}/12 formulas")
            if not perfect_labels:
                issues.append(f'all 4 category labels found in only {count_labels}/12 formulas')
            print(f"FAIL: Component 2 — {'; '.join(issues)}")

            # Award partial credit proportional to how many cells pass the label check
            # (most important: labels encode the classification logic)
            if count_labels > 0 and count_labels < 12:
                partial = round(0.4 * (count_labels / 12), 2)
                print(f"  Partial credit: {count_labels}/12 formulas have correct labels "
                      f"({partial} pts)")
                total_score += partial
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
