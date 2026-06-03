"""
Initial Setup: Contract urgency classification spreadsheet
Task ID: calc_fma_nested_if_date_050
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

WORKDIR = '/home/user'
TASK_ID = 'calc_fma_nested_if_date_050'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ContractStatus'

    # --- Headers ---
    ws['A1'] = 'ContractID'
    ws['B1'] = 'ExpiryDate'
    ws['C1'] = 'Urgency'

    # Style the header row
    header_font = Font(bold=True, name='Calibri', size=11)
    header_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}1']
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # --- Contract data ---
    # Dates selected to cover all four urgency categories relative to 2026-03-04:
    # Critical (<=30 days): expiry on 2026-03-10, 2026-03-20, 2026-03-31
    # Warning (<=90 days):  expiry on 2026-04-18, 2026-05-05, 2026-05-28
    # OK (<=365 days):      expiry on 2026-06-25, 2026-09-10, 2026-12-15
    # Long Term (>365 days): expiry on 2027-06-01, 2028-01-10, 2028-09-05
    contracts = [
        ('CTR-2026-001', date(2026, 3, 10)),   # Critical: 6 days away
        ('CTR-2026-002', date(2026, 3, 20)),   # Critical: 16 days away
        ('CTR-2026-003', date(2026, 3, 31)),   # Critical: 27 days away
        ('CTR-2026-004', date(2026, 4, 18)),   # Warning: 45 days away
        ('CTR-2026-005', date(2026, 5, 5)),    # Warning: 62 days away
        ('CTR-2026-006', date(2026, 5, 28)),   # Warning: 85 days away
        ('CTR-2026-007', date(2026, 6, 25)),   # OK: 113 days away
        ('CTR-2026-008', date(2026, 9, 10)),   # OK: 190 days away
        ('CTR-2026-009', date(2026, 12, 15)),  # OK: 286 days away
        ('CTR-2026-010', date(2027, 6, 1)),    # Long Term: 454 days away
        ('CTR-2026-011', date(2028, 1, 10)),   # Long Term: 677 days away
        ('CTR-2026-012', date(2028, 9, 5)),    # Long Term: 916 days away
    ]

    for row_idx, (contract_id, expiry_date) in enumerate(contracts, 2):
        ws.cell(row=row_idx, column=1, value=contract_id)
        date_cell = ws.cell(row=row_idx, column=2, value=expiry_date)
        date_cell.number_format = 'yyyy-mm-dd'
        # Column C (Urgency) is intentionally left empty

    # --- Column widths ---
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: ContractStatus')
    print(f'  Rows: 12 contract rows (rows 2-13)')
    print(f'  Column C (Urgency): EMPTY (task is to fill these with formulas)')


create_initial()
