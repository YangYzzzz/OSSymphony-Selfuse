"""
Initial Setup: HR Performance Rating Distribution
Task ID: calc_hr_performance_distribution_066
Domain: libreoffice_calc

Creates an Excel workbook with:
- Sheet 'Annual Reviews': 210 rows of employee annual review data with ratings 1-5
- Sheet 'Rating Distribution': exists but is completely empty (task requires filling it)
"""

import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_performance_distribution_066'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Annual Reviews ---
    ws1 = wb.active
    ws1.title = 'Annual Reviews'

    # Headers
    headers = ['Emp ID', 'Name', 'Department', 'Manager', 'Final Rating']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)
        ws1.cell(row=1, column=col).font = Font(bold=True)

    # Realistic employee data — 210 rows
    # Distribution roughly: 1=5%, 2=15%, 3=45%, 4=25%, 5=10%
    # That gives: 11 ones, 31 twos, 94 threes, 53 fours, 21 fives = 210 total
    employees = [
        # Emp ID, Name, Department, Manager
        ('E001', 'Sarah Chen', 'Engineering', 'David Park'),
        ('E002', 'Marcus Johnson', 'Marketing', 'Linda Torres'),
        ('E003', 'Emily Rodriguez', 'HR', 'Karen Mitchell'),
        ('E004', 'James Nguyen', 'Finance', 'Robert Walsh'),
        ('E005', 'Amanda Williams', 'Engineering', 'David Park'),
        ('E006', 'Tyler Brooks', 'Sales', 'Michael Reed'),
        ('E007', 'Priya Sharma', 'Engineering', 'David Park'),
        ('E008', 'Kevin O\'Brien', 'Marketing', 'Linda Torres'),
        ('E009', 'Diana Foster', 'Operations', 'Steve Coleman'),
        ('E010', 'Carlos Mendez', 'Finance', 'Robert Walsh'),
        ('E011', 'Rachel Kim', 'HR', 'Karen Mitchell'),
        ('E012', 'Nathan Scott', 'Engineering', 'David Park'),
        ('E013', 'Stephanie Lee', 'Sales', 'Michael Reed'),
        ('E014', 'Brian Murphy', 'Operations', 'Steve Coleman'),
        ('E015', 'Tiffany Jackson', 'Marketing', 'Linda Torres'),
        ('E016', 'Andrew Clark', 'Engineering', 'David Park'),
        ('E017', 'Jennifer White', 'Finance', 'Robert Walsh'),
        ('E018', 'Robert Hill', 'Sales', 'Michael Reed'),
        ('E019', 'Maria Garcia', 'HR', 'Karen Mitchell'),
        ('E020', 'Christopher Lewis', 'Engineering', 'David Park'),
        ('E021', 'Laura Martinez', 'Operations', 'Steve Coleman'),
        ('E022', 'Daniel Robinson', 'Marketing', 'Linda Torres'),
        ('E023', 'Ashley Walker', 'Finance', 'Robert Walsh'),
        ('E024', 'Jonathan Hall', 'Engineering', 'David Park'),
        ('E025', 'Samantha Allen', 'Sales', 'Michael Reed'),
        ('E026', 'Patrick Young', 'HR', 'Karen Mitchell'),
        ('E027', 'Melissa King', 'Engineering', 'David Park'),
        ('E028', 'Timothy Wright', 'Operations', 'Steve Coleman'),
        ('E029', 'Natalie Scott', 'Marketing', 'Linda Torres'),
        ('E030', 'Gregory Adams', 'Finance', 'Robert Walsh'),
        ('E031', 'Hannah Baker', 'Engineering', 'David Park'),
        ('E032', 'Steven Gonzalez', 'Sales', 'Michael Reed'),
        ('E033', 'Elizabeth Nelson', 'HR', 'Karen Mitchell'),
        ('E034', 'William Carter', 'Engineering', 'David Park'),
        ('E035', 'Jessica Mitchell', 'Operations', 'Steve Coleman'),
        ('E036', 'Ryan Perez', 'Marketing', 'Linda Torres'),
        ('E037', 'Nicole Roberts', 'Finance', 'Robert Walsh'),
        ('E038', 'Edward Turner', 'Engineering', 'David Park'),
        ('E039', 'Megan Phillips', 'Sales', 'Michael Reed'),
        ('E040', 'Joshua Campbell', 'HR', 'Karen Mitchell'),
        ('E041', 'Crystal Parker', 'Engineering', 'David Park'),
        ('E042', 'Brandon Evans', 'Operations', 'Steve Coleman'),
        ('E043', 'Vanessa Edwards', 'Marketing', 'Linda Torres'),
        ('E044', 'Zachary Collins', 'Finance', 'Robert Walsh'),
        ('E045', 'Angela Stewart', 'Engineering', 'David Park'),
        ('E046', 'Matthew Sanchez', 'Sales', 'Michael Reed'),
        ('E047', 'Michelle Morris', 'HR', 'Karen Mitchell'),
        ('E048', 'Aaron Rogers', 'Engineering', 'David Park'),
        ('E049', 'Kimberly Reed', 'Operations', 'Steve Coleman'),
        ('E050', 'Justin Cook', 'Marketing', 'Linda Torres'),
        ('E051', 'Heather Bailey', 'Finance', 'Robert Walsh'),
        ('E052', 'Sean Bell', 'Engineering', 'David Park'),
        ('E053', 'Amber Rivera', 'Sales', 'Michael Reed'),
        ('E054', 'Derrick Cooper', 'HR', 'Karen Mitchell'),
        ('E055', 'Courtney Richardson', 'Engineering', 'David Park'),
        ('E056', 'Travis Cox', 'Operations', 'Steve Coleman'),
        ('E057', 'Alicia Howard', 'Marketing', 'Linda Torres'),
        ('E058', 'Jared Ward', 'Finance', 'Robert Walsh'),
        ('E059', 'Whitney Torres', 'Engineering', 'David Park'),
        ('E060', 'Cody Peterson', 'Sales', 'Michael Reed'),
        ('E061', 'Brittany Gray', 'HR', 'Karen Mitchell'),
        ('E062', 'Chad Ramirez', 'Engineering', 'David Park'),
        ('E063', 'Monica James', 'Operations', 'Steve Coleman'),
        ('E064', 'Dustin Watson', 'Marketing', 'Linda Torres'),
        ('E065', 'Kayla Brooks', 'Finance', 'Robert Walsh'),
        ('E066', 'Phillip Kelly', 'Engineering', 'David Park'),
        ('E067', 'Sheila Sanders', 'Sales', 'Michael Reed'),
        ('E068', 'Lance Price', 'HR', 'Karen Mitchell'),
        ('E069', 'Tricia Bennett', 'Engineering', 'David Park'),
        ('E070', 'Rodney Wood', 'Operations', 'Steve Coleman'),
        ('E071', 'Felicia Barnes', 'Marketing', 'Linda Torres'),
        ('E072', 'Darren Ross', 'Finance', 'Robert Walsh'),
        ('E073', 'Jacqueline Henderson', 'Engineering', 'David Park'),
        ('E074', 'Leroy Coleman', 'Sales', 'Michael Reed'),
        ('E075', 'Sandra Jenkins', 'HR', 'Karen Mitchell'),
        ('E076', 'Marvin Perry', 'Engineering', 'David Park'),
        ('E077', 'Tammie Powell', 'Operations', 'Steve Coleman'),
        ('E078', 'Cecil Long', 'Marketing', 'Linda Torres'),
        ('E079', 'Geraldine Patterson', 'Finance', 'Robert Walsh'),
        ('E080', 'Clarence Hughes', 'Engineering', 'David Park'),
        ('E081', 'Bernice Flores', 'Sales', 'Michael Reed'),
        ('E082', 'Frederick Washington', 'HR', 'Karen Mitchell'),
        ('E083', 'Ethel Butler', 'Engineering', 'David Park'),
        ('E084', 'Willie Simmons', 'Operations', 'Steve Coleman'),
        ('E085', 'Jessie Foster', 'Marketing', 'Linda Torres'),
        ('E086', 'Maxine Gonzales', 'Finance', 'Robert Walsh'),
        ('E087', 'Clifford Bryant', 'Engineering', 'David Park'),
        ('E088', 'Olga Alexander', 'Sales', 'Michael Reed'),
        ('E089', 'Victor Russell', 'HR', 'Karen Mitchell'),
        ('E090', 'Irene Griffin', 'Engineering', 'David Park'),
        ('E091', 'Lonnie Diaz', 'Operations', 'Steve Coleman'),
        ('E092', 'Dolores Hayes', 'Marketing', 'Linda Torres'),
        ('E093', 'Willard Myers', 'Finance', 'Robert Walsh'),
        ('E094', 'Gladys Ford', 'Engineering', 'David Park'),
        ('E095', 'Terrence Hamilton', 'Sales', 'Michael Reed'),
        ('E096', 'Norma Graham', 'HR', 'Karen Mitchell'),
        ('E097', 'Alton Sullivan', 'Engineering', 'David Park'),
        ('E098', 'Vera Wallace', 'Operations', 'Steve Coleman'),
        ('E099', 'Edmund Woods', 'Marketing', 'Linda Torres'),
        ('E100', 'Stella Cole', 'Finance', 'Robert Walsh'),
        ('E101', 'Floyd West', 'Engineering', 'David Park'),
        ('E102', 'Lena Jordan', 'Sales', 'Michael Reed'),
        ('E103', 'Homer Owens', 'HR', 'Karen Mitchell'),
        ('E104', 'Pauline Reynolds', 'Engineering', 'David Park'),
        ('E105', 'Grady Fisher', 'Operations', 'Steve Coleman'),
        ('E106', 'Minnie Ellis', 'Marketing', 'Linda Torres'),
        ('E107', 'Lloyd Harrison', 'Finance', 'Robert Walsh'),
        ('E108', 'Essie Gibson', 'Engineering', 'David Park'),
        ('E109', 'Amos Mcdonald', 'Sales', 'Michael Reed'),
        ('E110', 'Lillie Cruz', 'HR', 'Karen Mitchell'),
        ('E111', 'Clifton Marshall', 'Engineering', 'David Park'),
        ('E112', 'Fannie Ortiz', 'Operations', 'Steve Coleman'),
        ('E113', 'Herman Gomez', 'Marketing', 'Linda Torres'),
        ('E114', 'Leona Murray', 'Finance', 'Robert Walsh'),
        ('E115', 'Otis Freeman', 'Engineering', 'David Park'),
        ('E116', 'Hattie Wells', 'Sales', 'Michael Reed'),
        ('E117', 'Roosevelt Webb', 'HR', 'Karen Mitchell'),
        ('E118', 'Ida Simpson', 'Engineering', 'David Park'),
        ('E119', 'Clyde Stevens', 'Operations', 'Steve Coleman'),
        ('E120', 'Marguerite Tucker', 'Marketing', 'Linda Torres'),
        ('E121', 'Elmer Porter', 'Finance', 'Robert Walsh'),
        ('E122', 'Loretta Hunter', 'Engineering', 'David Park'),
        ('E123', 'Horace Hicks', 'Sales', 'Michael Reed'),
        ('E124', 'Celia Crawford', 'HR', 'Karen Mitchell'),
        ('E125', 'Buford Henry', 'Engineering', 'David Park'),
        ('E126', 'Mattie Boyd', 'Operations', 'Steve Coleman'),
        ('E127', 'Dewey Mason', 'Marketing', 'Linda Torres'),
        ('E128', 'Myrtle Morales', 'Finance', 'Robert Walsh'),
        ('E129', 'Rufus Kennedy', 'Engineering', 'David Park'),
        ('E130', 'Bessie Warren', 'Sales', 'Michael Reed'),
        ('E131', 'Virgil Dixon', 'HR', 'Karen Mitchell'),
        ('E132', 'Thelma Ramos', 'Engineering', 'David Park'),
        ('E133', 'Sylvester Burns', 'Operations', 'Steve Coleman'),
        ('E134', 'Mamie Gordon', 'Marketing', 'Linda Torres'),
        ('E135', 'Leroy Shaw', 'Finance', 'Robert Walsh'),
        ('E136', 'Mabel Holmes', 'Engineering', 'David Park'),
        ('E137', 'Dwayne Rice', 'Sales', 'Michael Reed'),
        ('E138', 'Audrey Robertson', 'HR', 'Karen Mitchell'),
        ('E139', 'Elijah Henderson', 'Engineering', 'David Park'),
        ('E140', 'Gertrude Hunt', 'Operations', 'Steve Coleman'),
        ('E141', 'Cornelius Black', 'Marketing', 'Linda Torres'),
        ('E142', 'Blanche Daniels', 'Finance', 'Robert Walsh'),
        ('E143', 'Columbus Palmer', 'Engineering', 'David Park'),
        ('E144', 'Ora Mills', 'Sales', 'Michael Reed'),
        ('E145', 'Archie Nichols', 'HR', 'Karen Mitchell'),
        ('E146', 'Lula Grant', 'Engineering', 'David Park'),
        ('E147', 'Reginald Knight', 'Operations', 'Steve Coleman'),
        ('E148', 'Nora Ferguson', 'Marketing', 'Linda Torres'),
        ('E149', 'Percival Rose', 'Finance', 'Robert Walsh'),
        ('E150', 'Edna Stone', 'Engineering', 'David Park'),
        ('E151', 'Hubert Hawkins', 'Sales', 'Michael Reed'),
        ('E152', 'Mildred Dunn', 'HR', 'Karen Mitchell'),
        ('E153', 'Lester Perkins', 'Engineering', 'David Park'),
        ('E154', 'Sadie Payne', 'Operations', 'Steve Coleman'),
        ('E155', 'Roscoe Pierce', 'Marketing', 'Linda Torres'),
        ('E156', 'Beulah Berry', 'Finance', 'Robert Walsh'),
        ('E157', 'Chester Webb', 'Engineering', 'David Park'),
        ('E158', 'Carrie Strickland', 'Sales', 'Michael Reed'),
        ('E159', 'Booker Wagner', 'HR', 'Karen Mitchell'),
        ('E160', 'Pearl Walton', 'Engineering', 'David Park'),
        ('E161', 'Harley Snyder', 'Operations', 'Steve Coleman'),
        ('E162', 'Susie Casey', 'Marketing', 'Linda Torres'),
        ('E163', 'Newton Bell', 'Finance', 'Robert Walsh'),
        ('E164', 'Cordelia Sims', 'Engineering', 'David Park'),
        ('E165', 'Lemuel Castillo', 'Sales', 'Michael Reed'),
        ('E166', 'Viola Flynn', 'HR', 'Karen Mitchell'),
        ('E167', 'Emmet Obrien', 'Engineering', 'David Park'),
        ('E168', 'Mable Walters', 'Operations', 'Steve Coleman'),
        ('E169', 'Delbert Fuller', 'Marketing', 'Linda Torres'),
        ('E170', 'Leila Williamson', 'Finance', 'Robert Walsh'),
        ('E171', 'Alonzo Lawson', 'Engineering', 'David Park'),
        ('E172', 'Geneva Jacobs', 'Sales', 'Michael Reed'),
        ('E173', 'Garfield Chavez', 'HR', 'Karen Mitchell'),
        ('E174', 'Josephine Garza', 'Engineering', 'David Park'),
        ('E175', 'Cornelius Wade', 'Operations', 'Steve Coleman'),
        ('E176', 'Harriet Barnett', 'Marketing', 'Linda Torres'),
        ('E177', 'Edmund Dunn', 'Finance', 'Robert Walsh'),
        ('E178', 'Maude Malone', 'Engineering', 'David Park'),
        ('E179', 'Isidore Vega', 'Sales', 'Michael Reed'),
        ('E180', 'Almeda Drake', 'HR', 'Karen Mitchell'),
        ('E181', 'Oswald Pena', 'Engineering', 'David Park'),
        ('E182', 'Constance Farmer', 'Operations', 'Steve Coleman'),
        ('E183', 'Wendell Hammond', 'Marketing', 'Linda Torres'),
        ('E184', 'Elvira Thornton', 'Finance', 'Robert Walsh'),
        ('E185', 'Jasper Mendoza', 'Engineering', 'David Park'),
        ('E186', 'Opal Burgess', 'Sales', 'Michael Reed'),
        ('E187', 'Forrest Cobb', 'HR', 'Karen Mitchell'),
        ('E188', 'Lavinia Santos', 'Engineering', 'David Park'),
        ('E189', 'Hosea Obrien', 'Operations', 'Steve Coleman'),
        ('E190', 'Magnolia Gutierrez', 'Marketing', 'Linda Torres'),
        ('E191', 'Judson Carr', 'Finance', 'Robert Walsh'),
        ('E192', 'Octavia Barker', 'Engineering', 'David Park'),
        ('E193', 'Ezra Hoffman', 'Sales', 'Michael Reed'),
        ('E194', 'Celestine Swanson', 'HR', 'Karen Mitchell'),
        ('E195', 'Thaddeus Shelton', 'Engineering', 'David Park'),
        ('E196', 'Theodora Cannon', 'Operations', 'Steve Coleman'),
        ('E197', 'Barnabas Yates', 'Marketing', 'Linda Torres'),
        ('E198', 'Sophronia Frazier', 'Finance', 'Robert Walsh'),
        ('E199', 'Lucius Mckinney', 'Engineering', 'David Park'),
        ('E200', 'Arabella Calhoun', 'Sales', 'Michael Reed'),
        ('E201', 'Balthazar Weston', 'HR', 'Karen Mitchell'),
        ('E202', 'Seraphina Crosby', 'Engineering', 'David Park'),
        ('E203', 'Thaddeus Burke', 'Operations', 'Steve Coleman'),
        ('E204', 'Millicent Nolan', 'Marketing', 'Linda Torres'),
        ('E205', 'Barnabas Figueroa', 'Finance', 'Robert Walsh'),
        ('E206', 'Sophronia Klein', 'Engineering', 'David Park'),
        ('E207', 'Lucius Sparks', 'Sales', 'Michael Reed'),
        ('E208', 'Arabella Townsend', 'HR', 'Karen Mitchell'),
        ('E209', 'Ezekiel Bowers', 'Engineering', 'David Park'),
        ('E210', 'Theophilus Gallagher', 'Operations', 'Steve Coleman'),
    ]

    # Assign ratings: distribution roughly 5%, 15%, 45%, 25%, 10%
    # Rating 1: 11 employees (rows 2-12)
    # Rating 2: 31 employees (rows 13-43)
    # Rating 3: 94 employees (rows 44-137)
    # Rating 4: 53 employees (rows 138-190)
    # Rating 5: 21 employees (rows 191-211)
    rating_1_count = 11
    rating_2_count = 31
    rating_3_count = 94
    rating_4_count = 53
    rating_5_count = 21

    ratings = (
        [1] * rating_1_count +
        [2] * rating_2_count +
        [3] * rating_3_count +
        [4] * rating_4_count +
        [5] * rating_5_count
    )

    for i, (emp_data, rating) in enumerate(zip(employees, ratings)):
        row = i + 2
        emp_id, name, dept, manager = emp_data
        ws1.cell(row=row, column=1, value=emp_id)
        ws1.cell(row=row, column=2, value=name)
        ws1.cell(row=row, column=3, value=dept)
        ws1.cell(row=row, column=4, value=manager)
        ws1.cell(row=row, column=5, value=rating)

    # --- Sheet 2: Rating Distribution (empty — task requires filling this) ---
    ws2 = wb.create_sheet('Rating Distribution')
    # Leave completely empty — no headers, no data, no formulas

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Annual Reviews: 210 data rows with ratings 1-5')
    print(f'  Rating 1: {rating_1_count}, Rating 2: {rating_2_count}, '
          f'Rating 3: {rating_3_count}, Rating 4: {rating_4_count}, '
          f'Rating 5: {rating_5_count}')
    print(f'  Rating Distribution: empty sheet (for agent to fill)')


create_initial()
