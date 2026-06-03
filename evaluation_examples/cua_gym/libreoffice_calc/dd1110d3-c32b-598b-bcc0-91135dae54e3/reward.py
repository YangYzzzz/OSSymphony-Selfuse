"""
Reward Script: HR Performance Rating Distribution
Task ID: calc_hr_performance_distribution_066
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20): Headers present and bold in Rating Distribution sheet
  Component 2 (0.30): Rating values (1-5) and COUNTIF formulas in B2:B6
  Component 3 (0.20): Percentage formulas in C2:C6 (0.0% format) + Total row with SUM formula (bold)
  Component 4 (0.30): Bar/column chart present with title 'Performance Rating Distribution'
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_performance_distribution_066'


def extract_chart_title(chart):
    """Extract chart title text from openpyxl chart object."""
    try:
        for p in chart.title.tx.rich.p:
            for r in p.r:
                return r.t
    except Exception:
        pass
    try:
        return str(chart.title)
    except Exception:
        pass
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: 'Rating Distribution' sheet must exist
    if 'Rating Distribution' not in wb.sheetnames:
        print("FAIL: Sheet 'Rating Distribution' not found in workbook")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws = wb['Rating Distribution']

    # Component 1: Headers present with bold formatting (0.20 points)
    # Expected: A1='Rating' (bold), B1='Count' (bold), C1='Percentage' (bold)
    # This FAILS on initial (empty sheet) -> PASSES on golden
    try:
        a1_val = ws['A1'].value
        b1_val = ws['B1'].value
        c1_val = ws['C1'].value
        a1_bold = ws['A1'].font.bold
        b1_bold = ws['B1'].font.bold
        c1_bold = ws['C1'].font.bold

        headers_correct = (
            str(a1_val).strip() == 'Rating' and
            str(b1_val).strip() == 'Count' and
            str(c1_val).strip() == 'Percentage'
        )
        headers_bold = (a1_bold and b1_bold and c1_bold)

        if headers_correct and headers_bold:
            print(f"PASS: Component 1 — Headers A1='Rating', B1='Count', C1='Percentage' present and bold (0.20 pts)")
            total_score += 0.20
        elif headers_correct:
            print(f"FAIL: Component 1 — Headers present but not all bold (A1_bold={a1_bold}, B1_bold={b1_bold}, C1_bold={c1_bold})")
        else:
            print(f"FAIL: Component 1 — Expected headers 'Rating'/'Count'/'Percentage', found: A1={repr(a1_val)}, B1={repr(b1_val)}, C1={repr(c1_val)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Rating values 1-5 in A2:A6 and COUNTIF formulas in B2:B6 (0.30 points)
    # This FAILS on initial (empty sheet) -> PASSES on golden
    try:
        rating_failures = []
        countif_failures = []

        for row in range(2, 7):
            expected_rating = row - 1  # 1,2,3,4,5
            a_val = ws.cell(row=row, column=1).value
            b_val = ws.cell(row=row, column=2).value

            # Check rating value in column A
            if a_val != expected_rating:
                rating_failures.append(f"A{row}={repr(a_val)}, expected {expected_rating}")

            # Check COUNTIF formula in column B
            if not isinstance(b_val, str) or 'COUNTIF' not in b_val.upper():
                countif_failures.append(f"B{row}={repr(b_val)}")

        if not rating_failures and not countif_failures:
            print(f"PASS: Component 2 — Rating values 1-5 in A2:A6 and COUNTIF formulas in B2:B6 (0.30 pts)")
            total_score += 0.30
        else:
            if rating_failures:
                print(f"FAIL: Component 2 — Rating values incorrect: {rating_failures}")
            if countif_failures:
                print(f"FAIL: Component 2 — COUNTIF formulas missing/incorrect: {countif_failures}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Percentage formulas in C2:C6 (0.0% format) + Total row (A7='Total', B7=SUM, B7 bold) (0.20 points)
    # This FAILS on initial (empty sheet) -> PASSES on golden
    try:
        pct_failures = []
        pct_fmt_failures = []
        total_row_failures = []

        # Check C2:C6 percentage formulas
        for row in range(2, 7):
            c_val = ws.cell(row=row, column=3).value
            c_fmt = ws.cell(row=row, column=3).number_format

            if not isinstance(c_val, str) or '/' not in c_val:
                pct_failures.append(f"C{row}={repr(c_val)}")
            elif '$B$7' not in c_val:
                pct_failures.append(f"C{row}={repr(c_val)} (missing $B$7 reference)")

            # Check number format is percentage-like (0.0% or similar)
            if not (c_fmt and '0' in c_fmt and '%' in c_fmt):
                pct_fmt_failures.append(f"C{row} format={repr(c_fmt)}")

        # Check Total row: A7='Total', B7=SUM formula, B7 bold
        a7_val = ws.cell(row=7, column=1).value
        b7_val = ws.cell(row=7, column=2).value
        b7_bold = ws.cell(row=7, column=2).font.bold

        if str(a7_val).strip() != 'Total':
            total_row_failures.append(f"A7={repr(a7_val)}, expected 'Total'")

        if not isinstance(b7_val, str) or 'SUM' not in b7_val.upper():
            total_row_failures.append(f"B7={repr(b7_val)}, expected SUM formula")

        if not b7_bold:
            total_row_failures.append(f"B7 not bold")

        all_failures = pct_failures + pct_fmt_failures + total_row_failures
        if not all_failures:
            print(f"PASS: Component 3 — Percentage formulas (C2:C6 with 0.0% format, referencing $B$7) and Total row (A7='Total', B7=SUM bold) (0.20 pts)")
            total_score += 0.20
        else:
            for fail in all_failures:
                print(f"FAIL: Component 3 — {fail}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Bar/column chart present with title 'Performance Rating Distribution' (0.30 points)
    # This FAILS on initial (empty sheet, no charts) -> PASSES on golden
    try:
        charts = ws._charts
        if len(charts) == 0:
            print(f"FAIL: Component 4 — No chart found in 'Rating Distribution' sheet")
        else:
            chart = charts[0]
            # Check it's a bar/column chart
            is_bar_chart = hasattr(chart, 'type') and (chart.type in ('col', 'bar'))
            chart_title = extract_chart_title(chart)

            if not is_bar_chart:
                print(f"FAIL: Component 4 — Chart should be a bar/column chart, found type={getattr(chart, 'type', 'unknown')}, class={chart.__class__.__name__}")
            elif chart_title and 'Performance Rating Distribution' in chart_title:
                print(f"PASS: Component 4 — Bar/column chart with title '{chart_title}' found (0.30 pts)")
                total_score += 0.30
            else:
                print(f"FAIL: Component 4 — Chart type correct ({chart.type}) but title incorrect: expected 'Performance Rating Distribution', found {repr(chart_title)}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
