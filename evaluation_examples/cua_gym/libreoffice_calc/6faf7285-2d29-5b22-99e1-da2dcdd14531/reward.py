"""
Reward Script: Create 'HighlightTotal' cell style and apply to totals row
Task ID: calc_gg3_044
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): Bold font on A12:D12
  Component 2 (0.35): Light green background (#90EE90) on A12:D12
  Component 3 (0.30): Thick bottom border on A12:D12
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_044'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Totals' sheet must exist
    if 'Totals' not in wb.sheetnames:
        print(f"FAIL: 'Totals' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Totals']
    target_cells = ['A12', 'B12', 'C12', 'D12']

    # Component 1: Bold font on A12:D12 (0.35 points)
    # Initial state: bold=False on all four cells
    # Golden state: bold=True on all four cells
    try:
        bold_count = 0
        for coord in target_cells:
            cell = ws[coord]
            if cell.font and cell.font.bold:
                bold_count += 1
            else:
                print(f"  DETAIL: {coord} bold={cell.font.bold if cell.font else None}")

        if bold_count == 4:
            print(f"PASS: Component 1 — All 4 cells (A12:D12) have bold font (0.35 pts)")
            total_score += 0.35
        elif bold_count > 0:
            partial = round(0.35 * (bold_count / 4), 2)
            print(f"PARTIAL: Component 1 — {bold_count}/4 cells have bold font ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No cells in A12:D12 have bold font")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Light green background (#90EE90 -> ARGB FF90EE90) on A12:D12 (0.35 points)
    # Initial state: fill=00000000 (no fill) on all four cells
    # Golden state: fill=FF90EE90 on all four cells
    try:
        green_count = 0
        expected_color = 'FF90EE90'
        for coord in target_cells:
            cell = ws[coord]
            try:
                fg_rgb = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
                if fg_rgb and fg_rgb.upper() == expected_color:
                    green_count += 1
                else:
                    print(f"  DETAIL: {coord} fill fgColor={fg_rgb}")
            except Exception:
                print(f"  DETAIL: {coord} fill check failed")

        if green_count == 4:
            print(f"PASS: Component 2 — All 4 cells (A12:D12) have light green background {expected_color} (0.35 pts)")
            total_score += 0.35
        elif green_count > 0:
            partial = round(0.35 * (green_count / 4), 2)
            print(f"PARTIAL: Component 2 — {green_count}/4 cells have green background ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No cells in A12:D12 have light green background (expected {expected_color})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Thick bottom border on A12:D12 (0.30 points)
    # Initial state: bottom border=None on all four cells
    # Golden state: bottom border=thick on all four cells
    try:
        border_count = 0
        for coord in target_cells:
            cell = ws[coord]
            if cell.border and cell.border.bottom and cell.border.bottom.style:
                # Accept 'thick' or 'medium' as valid thick-ish borders
                if cell.border.bottom.style in ('thick', 'medium'):
                    border_count += 1
                else:
                    print(f"  DETAIL: {coord} bottom border style={cell.border.bottom.style} (expected thick)")
            else:
                print(f"  DETAIL: {coord} no bottom border")

        if border_count == 4:
            print(f"PASS: Component 3 — All 4 cells (A12:D12) have thick bottom border (0.30 pts)")
            total_score += 0.30
        elif border_count > 0:
            partial = round(0.30 * (border_count / 4), 2)
            print(f"PARTIAL: Component 3 — {border_count}/4 cells have thick bottom border ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No cells in A12:D12 have thick bottom border")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved LibreOffice edits before verification
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(1.0)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
persist_app_state()

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
