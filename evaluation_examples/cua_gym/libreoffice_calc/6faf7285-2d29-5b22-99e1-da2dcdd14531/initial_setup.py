"""
Initial Setup: Create a budget tracker spreadsheet with a Totals sheet
Task ID: calc_gg3_044
Domain: libreoffice_calc

The Totals sheet has a budget table with line items in rows 2-11 and
total values in row 12. Row 12 has NO special formatting (plain like data rows).
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_044'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Totals sheet ---
    ws = wb.active
    ws.title = 'Totals'

    headers = ['Category', 'Budget', 'Actual', 'Variance']
    # Light header styling (just bold text, no color - plain look)
    header_font = Font(name='Calibri', size=11, bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    # Budget line items (rows 2-11) - realistic department budget data
    data = [
        ['Personnel',          125000.00, 118450.00,  6550.00],
        ['Office Supplies',      8500.00,   9120.00,  -620.00],
        ['IT Infrastructure',   45000.00,  43200.00,  1800.00],
        ['Marketing',           32000.00,  35600.00, -3600.00],
        ['Travel & Events',     18000.00,  16750.00,  1250.00],
        ['Professional Services',22000.00,  24100.00, -2100.00],
        ['Facilities',          15500.00,  15500.00,     0.00],
        ['Training & Development',9800.00,   8400.00,  1400.00],
        ['Insurance',           12000.00,  12000.00,     0.00],
        ['Miscellaneous',        5200.00,   6380.00, -1180.00],
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c + 1, value=row_data[c])
            cell.number_format = '#,##0.00'

    # Totals row (row 12) - plain formatting, same as data rows
    totals = [
        'Total',
        sum(d[1] for d in data),
        sum(d[2] for d in data),
        sum(d[3] for d in data),
    ]
    ws.cell(row=12, column=1, value=totals[0])
    for c in range(1, 4):
        cell = ws.cell(row=12, column=c + 1, value=totals[c])
        cell.number_format = '#,##0.00'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
