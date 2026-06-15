"""
Initial Setup: Format numbers in millions with 'M' suffix
Task ID: calc_lf_077
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_077'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Finance ---
    ws = wb.active
    ws.title = 'Finance'

    # Headers
    ws.cell(row=1, column=1, value='Metric')
    ws.cell(row=1, column=2, value='Value')

    # Header styling
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    for col in [1, 2]:
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")

    # Data rows - core task data
    data = [
        ['Revenue', 2500000],
        ['Assets', 15800000],
        ['Liability', 7200000],
    ]
    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        ws.cell(row=r, column=2, value=row_data[1])

    # Additional financial data for realism and complexity
    extra_data = [
        ['Net Income', 3400000],
        ['Operating Expenses', 9100000],
        ['Cash Flow', 1750000],
        ['Equity', 8600000],
        ['EBITDA', 4200000],
        ['Depreciation', 1300000],
        ['R&D Spending', 2800000],
        ['Marketing Budget', 1950000],
        ['Capital Expenditure', 3100000],
    ]
    for r, row_data in enumerate(extra_data, 5):
        ws.cell(row=r, column=1, value=row_data[0])
        ws.cell(row=r, column=2, value=row_data[1])

    # B column uses General format (no special formatting) - numbers display as plain integers
    # Do NOT apply the '#,##0.0,,"M"' format -- that is the task

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18

    # --- Sheet: Summary ---
    ws2 = wb.create_sheet('Summary')
    ws2.cell(row=1, column=1, value='Category')
    ws2.cell(row=1, column=2, value='Total')
    ws2.cell(row=2, column=1, value='All Metrics')
    ws2.cell(row=2, column=2, value='=SUM(Finance!B2:B13)')

    # --- Sheet: Notes ---
    ws3 = wb.create_sheet('Notes')
    ws3.cell(row=1, column=1, value='Financial Report Notes')
    ws3.cell(row=2, column=1, value='Report Period: Q1 2025')
    ws3.cell(row=3, column=1, value='Prepared by: Finance Department')
    ws3.cell(row=4, column=1, value='All values are in raw numbers. Apply appropriate display formatting as needed.')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
