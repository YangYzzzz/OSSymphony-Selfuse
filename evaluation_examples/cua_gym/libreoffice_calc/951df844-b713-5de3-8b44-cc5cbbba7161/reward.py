"""
Reward Script: Format numbers in millions with 'M' suffix
Task ID: calc_lf_077
Domain: libreoffice_calc
Scoring:
  - Component 1: B2 number format is '#,##0.0,,"M"' (0.35 pts)
  - Component 2: B3 number format is '#,##0.0,,"M"' (0.35 pts)
  - Component 3: B4 number format is '#,##0.0,,"M"' (0.30 pts)
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_077'
EXPECTED_FORMAT = '#,##0.0,,"M"'

def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify that cells B2:B4 on the 'Finance' sheet have the custom
    number format '#,##0.0,,"M"' applied.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the 'Finance' sheet exists (precondition gate)
    if 'Finance' not in wb.sheetnames:
        print("CRITICAL: 'Finance' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Finance']

    # Component 1: B2 has the millions format (0.35 points)
    try:
        nf = ws['B2'].number_format
        if nf == EXPECTED_FORMAT:
            print(f"PASS: Component 1 — B2 number_format is '{nf}' (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 1 — B2 number_format is '{nf}', expected '{EXPECTED_FORMAT}'")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: B3 has the millions format (0.35 points)
    try:
        nf = ws['B3'].number_format
        if nf == EXPECTED_FORMAT:
            print(f"PASS: Component 2 — B3 number_format is '{nf}' (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 2 — B3 number_format is '{nf}', expected '{EXPECTED_FORMAT}'")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: B4 has the millions format (0.30 points)
    try:
        nf = ws['B4'].number_format
        if nf == EXPECTED_FORMAT:
            print(f"PASS: Component 3 — B4 number_format is '{nf}' (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 3 — B4 number_format is '{nf}', expected '{EXPECTED_FORMAT}'")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist any unsaved GUI state before verification
persist_app_state("libreoffice_calc")

# Run verification
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
