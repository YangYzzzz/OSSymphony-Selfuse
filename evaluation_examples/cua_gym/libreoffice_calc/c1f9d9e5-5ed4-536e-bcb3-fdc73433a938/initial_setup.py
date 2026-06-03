"""
Initial Setup: Insert five new sheets before 'Summary' sheet
Task ID: calc_gsi_036
Domain: libreoffice_calc

Creates a workbook with a single 'Summary' sheet containing realistic
annual budget summary data. The agent must insert 5 monthly sheets
(Jan-May) before this sheet.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_036'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Summary Sheet ---
    ws = wb.active
    ws.title = 'Summary'

    # Header row styling
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    # Headers
    headers = ['Category', 'Q1 Budget', 'Q2 Budget', 'Q3 Budget', 'Q4 Budget', 'Annual Total']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Realistic budget data
    data = [
        ['Salaries & Wages', 245000, 248500, 252000, 255500, 1001000],
        ['Employee Benefits', 73500, 74550, 75600, 76650, 300300],
        ['Office Rent', 36000, 36000, 36000, 36000, 144000],
        ['Utilities', 8200, 7800, 9500, 8900, 34400],
        ['Software Licenses', 15600, 15600, 15600, 15600, 62400],
        ['Marketing & Advertising', 42000, 55000, 38000, 65000, 200000],
        ['Travel & Entertainment', 12500, 18000, 14000, 22000, 66500],
        ['Professional Services', 25000, 30000, 25000, 35000, 115000],
        ['Equipment & Supplies', 8500, 6200, 9800, 11500, 36000],
        ['Training & Development', 5000, 8000, 5000, 12000, 30000],
        ['Insurance', 18000, 18000, 18000, 18000, 72000],
        ['Depreciation', 15000, 15000, 15000, 15000, 60000],
        ['Contingency Fund', 10000, 10000, 10000, 10000, 40000],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c >= 2:
                cell.number_format = '$#,##0'

    # Totals row
    total_row = len(data) + 2
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    for col in range(2, 7):
        cell = ws.cell(row=total_row, column=col)
        cell.value = f'=SUM({chr(64+col)}2:{chr(64+col)}{total_row-1})'
        cell.font = Font(bold=True)
        cell.number_format = '$#,##0'

    # Column widths
    ws.column_dimensions['A'].width = 28
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col_letter].width = 15

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
