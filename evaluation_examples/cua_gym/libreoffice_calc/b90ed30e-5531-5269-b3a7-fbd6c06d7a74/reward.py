"""
Reward Script: Merge and format invoice header cell A1:G1
Task ID: calc_gsd_007
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30) - A1:G1 is merged into one cell
  Component 2 (0.15) - Merged cell contains 'Acme Corp Solutions'
  Component 3 (0.20) - Font size is 18pt and bold
  Component 4 (0.15) - Text is horizontally centered
  Component 5 (0.20) - Thick bottom border applied
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_007'


def persist_app_state(domain: str):
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Component 1: A1:G1 is merged (0.30 points)
    try:
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        # Check that A1:G1 is among the merged ranges
        a1g1_merged = any(mr.upper() == 'A1:G1' for mr in merged_ranges)
        if a1g1_merged:
            # Double-check: B1 through G1 should be MergedCell type
            all_merged = all(isinstance(ws.cell(row=1, column=c), MergedCell) for c in range(2, 8))
            if all_merged:
                print(f"PASS: Component 1 - A1:G1 is merged (0.30 pts)")
                total_score += 0.30
            else:
                print(f"FAIL: Component 1 - Merge range found but cells not all MergedCell type")
        else:
            print(f"FAIL: Component 1 - A1:G1 not in merged ranges. Found: {merged_ranges}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Merged cell contains 'Acme Corp Solutions' (0.15 points)
    # This is a compound check: cell must be merged AND have correct text
    try:
        a1_val = ws['A1'].value
        if a1_val is not None and str(a1_val).strip() == 'Acme Corp Solutions' and total_score >= 0.30:
            print(f"PASS: Component 2 - A1 contains 'Acme Corp Solutions' in merged cell (0.15 pts)")
            total_score += 0.15
        elif a1_val is not None and str(a1_val).strip() == 'Acme Corp Solutions':
            # Text is correct but cell is not merged - no credit (merge is prerequisite)
            print(f"FAIL: Component 2 - Text correct but cell not merged")
        else:
            print(f"FAIL: Component 2 - Expected 'Acme Corp Solutions', found: {repr(a1_val)}")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Font size 18pt and bold (0.20 points)
    try:
        font = ws['A1'].font
        is_bold = font.bold == True
        is_size_18 = font.size is not None and abs(float(font.size) - 18.0) < 0.5
        if is_bold and is_size_18:
            print(f"PASS: Component 3 - Font is 18pt bold (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 3 - bold={font.bold}, size={font.size} (expected bold=True, size=18)")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Horizontally centered (0.15 points)
    try:
        alignment = ws['A1'].alignment
        if alignment.horizontal == 'center':
            print(f"PASS: Component 4 - Text is horizontally centered (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 - horizontal alignment={repr(alignment.horizontal)}, expected 'center'")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Thick bottom border (0.20 points)
    # openpyxl 'thick' style corresponds to approx 2.5-3pt line
    try:
        border = ws['A1'].border
        bottom_style = border.bottom.style if border.bottom else None
        # 'thick' and 'medium' both qualify as thick borders (>=2.5pt)
        # openpyxl thick = ~3pt, medium = ~2pt
        # Task says >=2.5pt, so 'thick' is the expected match
        if bottom_style in ('thick',):
            print(f"PASS: Component 5 - Thick bottom border applied, style='{bottom_style}' (0.20 pts)")
            total_score += 0.20
        elif bottom_style in ('medium',):
            # Medium is ~2pt, close but not quite >=2.5pt; give partial
            print(f"PARTIAL: Component 5 - Medium bottom border found (partial: 0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 - bottom border style={repr(bottom_style)}, expected 'thick'")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
