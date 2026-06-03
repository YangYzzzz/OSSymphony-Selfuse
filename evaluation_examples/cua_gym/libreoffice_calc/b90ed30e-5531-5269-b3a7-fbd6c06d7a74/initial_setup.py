"""
Initial Setup: Invoice template with company name in A1, no formatting
Task ID: calc_gsd_007
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Invoice"

    # A1: Company name - plain text, no merge, no bold, no special formatting
    ws["A1"] = "Acme Corp Solutions"

    # Row 2: Invoice headers
    headers = ["Invoice #", "Date", "Due Date", "Description", "Qty", "Unit Price", "Total"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center")

    # Rows 3-20: Realistic invoice line items
    invoice_data = [
        ["INV-2025-0041", "2025-03-01", "2025-03-31", "Web Development - Phase 1", 1, 4500.00, 4500.00],
        ["INV-2025-0041", "2025-03-01", "2025-03-31", "UI/UX Design Services", 3, 1200.00, 3600.00],
        ["INV-2025-0041", "2025-03-01", "2025-03-31", "Database Migration", 1, 2800.00, 2800.00],
        ["INV-2025-0041", "2025-03-01", "2025-03-31", "Server Configuration", 2, 750.00, 1500.00],
        ["INV-2025-0041", "2025-03-01", "2025-03-31", "SSL Certificate Setup", 1, 350.00, 350.00],
        ["INV-2025-0041", "2025-03-01", "2025-03-31", "API Integration", 4, 980.00, 3920.00],
        ["INV-2025-0041", "2025-03-01", "2025-03-31", "QA Testing - Sprint 1", 2, 600.00, 1200.00],
        ["INV-2025-0041", "2025-03-01", "2025-03-31", "Cloud Hosting (Monthly)", 1, 450.00, 450.00],
        ["INV-2025-0041", "2025-03-01", "2025-03-31", "Technical Documentation", 1, 1100.00, 1100.00],
        ["INV-2025-0041", "2025-03-01", "2025-03-31", "Project Management", 5, 500.00, 2500.00],
        ["INV-2025-0041", "2025-03-01", "2025-03-31", "Security Audit", 1, 3200.00, 3200.00],
        ["INV-2025-0041", "2025-03-01", "2025-03-31", "Performance Optimization", 2, 875.00, 1750.00],
        ["INV-2025-0041", "2025-03-01", "2025-03-31", "Email System Migration", 1, 1600.00, 1600.00],
        ["INV-2025-0041", "2025-03-01", "2025-03-31", "Training Sessions", 3, 400.00, 1200.00],
        ["INV-2025-0041", "2025-03-01", "2025-03-31", "Backup Solution Setup", 1, 900.00, 900.00],
        ["INV-2025-0041", "2025-03-01", "2025-03-31", "Mobile App Prototype", 1, 5500.00, 5500.00],
        ["INV-2025-0041", "2025-03-01", "2025-03-31", "CRM Integration", 2, 1350.00, 2700.00],
        ["INV-2025-0041", "2025-03-01", "2025-03-31", "Analytics Dashboard", 1, 2200.00, 2200.00],
    ]

    for r, row_data in enumerate(invoice_data, 3):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14

    # Format currency columns
    for row in range(3, 21):
        ws.cell(row=row, column=6).number_format = '$#,##0.00'
        ws.cell(row=row, column=7).number_format = '$#,##0.00'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
