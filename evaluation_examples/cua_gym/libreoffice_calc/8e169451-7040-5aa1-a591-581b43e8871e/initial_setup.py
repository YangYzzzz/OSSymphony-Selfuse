"""
Initial Setup: Sales Revenue Forecast - FORECAST function and confidence band
Task ID: calc_sales_revenue_forecast_014
Domain: libreoffice_calc

Creates a spreadsheet with QuarterlyRevenue sheet containing:
- 8 quarters of historical revenue data (Q1 2023 - Q4 2024)
- Quarter 9 (Q1 2025) placeholder row with empty B10 (forecast target)
- Existing line chart plotting historical data A2:B9
- NO forecast formula, NO confidence band columns yet
"""

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_revenue_forecast_014'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: QuarterlyRevenue ---
    ws = wb.active
    ws.title = 'QuarterlyRevenue'

    # Headers
    ws['A1'] = 'Quarter'
    ws['B1'] = 'Revenue'

    # Style headers bold
    bold_font = Font(bold=True, size=12)
    ws['A1'].font = bold_font
    ws['B1'].font = bold_font

    # Historical data: Q1 2023 - Q4 2024 (8 quarters, numeric 1-8)
    # Revenue in actual dollar values
    historical_data = [
        (1, 3200000),   # Q1 2023 - $3.2M
        (2, 3800000),   # Q2 2023 - $3.8M
        (3, 4100000),   # Q3 2023 - $4.1M
        (4, 4600000),   # Q4 2023 - $4.6M
        (5, 4900000),   # Q1 2024 - $4.9M
        (6, 5300000),   # Q2 2024 - $5.3M
        (7, 5800000),   # Q3 2024 - $5.8M
        (8, 6200000),   # Q4 2024 - $6.2M
    ]

    for row_idx, (quarter, revenue) in enumerate(historical_data, 2):
        ws.cell(row=row_idx, column=1, value=quarter)
        ws.cell(row=row_idx, column=2, value=revenue)

    # Row 10: Q1 2025 (quarter 9) — forecast target
    ws['A10'] = 9   # Q1 2025 numeric quarter
    # B10 intentionally left EMPTY — this is where FORECAST formula goes

    # Quarter labels in column C for human readability (optional display)
    quarter_labels = [
        'Q1 2023', 'Q2 2023', 'Q3 2023', 'Q4 2023',
        'Q1 2024', 'Q2 2024', 'Q3 2024', 'Q4 2024',
    ]
    ws['C1'] = 'Quarter Label'
    ws['C1'].font = bold_font
    for idx, label in enumerate(quarter_labels, 2):
        ws.cell(row=idx, column=3, value=label)
    ws['C10'] = 'Q1 2025'

    # Number formatting for revenue column
    for row_idx in range(2, 11):
        ws.cell(row=row_idx, column=2).number_format = '$#,##0'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 14

    # Add header border
    thin = Side(style='thin', color='000000')
    for col in range(1, 4):
        cell = ws.cell(row=1, column=col)
        cell.border = Border(bottom=thin)

    # --- Existing Line Chart: Historical Revenue (A2:B9 only) ---
    line_chart = LineChart()
    line_chart.title = 'Quarterly Revenue Trend'
    line_chart.y_axis.title = 'Revenue ($)'
    line_chart.x_axis.title = 'Quarter'
    line_chart.style = 10

    # Revenue data series (B2:B9 - historical only, NOT including row 10)
    revenue_data = Reference(ws, min_col=2, min_row=1, max_row=9)
    line_chart.add_data(revenue_data, titles_from_data=True)

    # Categories from A2:A9 (quarters 1-8)
    categories = Reference(ws, min_col=1, min_row=2, max_row=9)
    line_chart.set_categories(categories)

    # Chart dimensions and position
    line_chart.width = 20
    line_chart.height = 12

    # Add chart to sheet
    ws.add_chart(line_chart, 'E2')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Contents:')
    print('  Sheet: QuarterlyRevenue')
    print('  A1: Quarter, B1: Revenue, C1: Quarter Label')
    print('  A2:A9 = 1-8 (numeric quarters), B2:B9 = historical revenue')
    print('  A10 = 9 (Q1 2025), B10 = EMPTY (forecast target)')
    print('  C2:C9 = quarter labels, C10 = Q1 2025')
    print('  Existing line chart: historical data A2:B9 at E2')
    print('  NO forecast formula, NO Upper/Lower Bound columns')


create_initial()
