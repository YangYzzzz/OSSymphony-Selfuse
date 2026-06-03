"""
Reward Script: Use FORECAST function to project next quarter revenue and add 90% confidence band
Task ID: calc_sales_revenue_forecast_014
Domain: libreoffice_calc
Scoring:
  - Component 1: B10 FORECAST formula (0.35 pts)
  - Component 2: C1='Upper Bound', D1='Lower Bound' headers (0.20 pts)
  - Component 3: C10 upper bound formula with 1.645*STDEV (0.20 pts)
  - Component 4: D10 lower bound formula with 1.645*STDEV (0.15 pts)
  - Component 5: Chart series extended to include row 10 (0.10 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_revenue_forecast_014'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the sheet exists
    if 'QuarterlyRevenue' not in wb.sheetnames:
        print("CRITICAL: Sheet 'QuarterlyRevenue' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['QuarterlyRevenue']

    # Component 1: B10 contains FORECAST formula referencing A10, B2:B9, A2:A9 (0.35 points)
    # This FAILS on initial (B10 is None) and PASSES on golden (has FORECAST formula)
    try:
        b10_val = ws['B10'].value
        if b10_val is None:
            print(f"FAIL: Component 1 — B10 is empty, expected FORECAST formula")
        elif not isinstance(b10_val, str):
            print(f"FAIL: Component 1 — B10 is not a formula: {repr(b10_val)}")
        else:
            b10_upper = b10_val.upper().replace(' ', '')
            # Check for FORECAST function with correct argument references
            if 'FORECAST(' in b10_upper and 'A10' in b10_upper and 'B2:B9' in b10_upper and 'A2:A9' in b10_upper:
                print(f"PASS: Component 1 — B10 contains FORECAST formula: {b10_val} (0.35 pts)")
                total_score += 0.35
            elif 'FORECAST' in b10_upper:
                # Partial: has FORECAST but with slightly different args
                print(f"PARTIAL: Component 1 — B10 has FORECAST but args may differ: {b10_val} (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 1 — B10 does not contain FORECAST formula: {repr(b10_val)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: C1='Upper Bound' and D1='Lower Bound' headers (0.20 points)
    # This FAILS on initial (C1='Quarter Label', D1=None) and PASSES on golden
    try:
        c1_val = ws['C1'].value
        d1_val = ws['D1'].value
        c1_ok = c1_val is not None and str(c1_val).strip().lower() == 'upper bound'
        d1_ok = d1_val is not None and str(d1_val).strip().lower() == 'lower bound'

        if c1_ok and d1_ok:
            print(f"PASS: Component 2 — C1='{c1_val}', D1='{d1_val}' (0.20 pts)")
            total_score += 0.20
        elif c1_ok:
            print(f"PARTIAL: Component 2 — C1='{c1_val}' OK, but D1='{d1_val}' is not 'Lower Bound' (0.10 pts)")
            total_score += 0.10
        elif d1_ok:
            print(f"PARTIAL: Component 2 — D1='{d1_val}' OK, but C1='{c1_val}' is not 'Upper Bound' (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 — C1='{c1_val}', D1='{d1_val}', expected 'Upper Bound' and 'Lower Bound'")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: C10 contains upper bound formula with 1.645*STDEV pattern (0.20 points)
    # This FAILS on initial (C10='Q1 2025') and PASSES on golden (=B10+1.645*STDEV(B2:B9))
    try:
        c10_val = ws['C10'].value
        if c10_val is None:
            print(f"FAIL: Component 3 — C10 is empty, expected upper bound formula")
        elif not isinstance(c10_val, str):
            print(f"FAIL: Component 3 — C10 is not a formula: {repr(c10_val)}")
        else:
            c10_upper = c10_val.upper().replace(' ', '')
            has_b10 = 'B10' in c10_upper
            has_stdev = 'STDEV(' in c10_upper
            has_1645 = '1.645' in c10_upper
            # Check it's an addition (upper bound, not lower)
            has_plus = '+' in c10_val

            if has_b10 and has_stdev and has_1645 and has_plus:
                print(f"PASS: Component 3 — C10 upper bound formula correct: {c10_val} (0.20 pts)")
                total_score += 0.20
            elif has_stdev and has_1645:
                print(f"PARTIAL: Component 3 — C10 has STDEV*1.645 but may differ: {c10_val} (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 3 — C10 does not match expected upper bound pattern: {repr(c10_val)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: D10 contains lower bound formula with 1.645*STDEV pattern (0.15 points)
    # This FAILS on initial (D10=None) and PASSES on golden (=B10-1.645*STDEV(B2:B9))
    try:
        d10_val = ws['D10'].value
        if d10_val is None:
            print(f"FAIL: Component 4 — D10 is empty, expected lower bound formula")
        elif not isinstance(d10_val, str):
            print(f"FAIL: Component 4 — D10 is not a formula: {repr(d10_val)}")
        else:
            d10_upper = d10_val.upper().replace(' ', '')
            has_b10 = 'B10' in d10_upper
            has_stdev = 'STDEV(' in d10_upper
            has_1645 = '1.645' in d10_upper
            # Check it's a subtraction (lower bound)
            has_minus = '-' in d10_val

            if has_b10 and has_stdev and has_1645 and has_minus:
                print(f"PASS: Component 4 — D10 lower bound formula correct: {d10_val} (0.15 pts)")
                total_score += 0.15
            elif has_stdev and has_1645:
                print(f"PARTIAL: Component 4 — D10 has STDEV*1.645 but may differ: {d10_val} (0.07 pts)")
                total_score += 0.07
            else:
                print(f"FAIL: Component 4 — D10 does not match expected lower bound pattern: {repr(d10_val)}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Chart series data reference extended to include row 10 (0.10 points)
    # Initial chart references '$B$2:$B$9' (8 data points)
    # Golden chart references '$B$2:$B$10' (9 data points including forecast)
    try:
        charts = ws._charts if hasattr(ws, '_charts') else []
        if not charts:
            print("FAIL: Component 5 — No charts found on sheet")
        else:
            chart = charts[0]
            found_extended_ref = None
            for ser in chart.series:
                if hasattr(ser, 'val') and ser.val is not None:
                    val_ref = ser.val
                    if hasattr(val_ref, 'numRef') and val_ref.numRef is not None:
                        ref_formula = str(val_ref.numRef.f)
                        # Check if chart data now includes row 10
                        if 'B$10' in ref_formula or 'B10' in ref_formula or ':$B$10' in ref_formula:
                            found_extended_ref = ref_formula
                            break

            if found_extended_ref is not None:
                print(f"PASS: Component 5 — Chart series extended to include row 10: {found_extended_ref} (0.10 pts)")
                total_score += 0.10
            else:
                # Get the actual reference for diagnostic info
                ref_info = 'unknown'
                try:
                    ref_info = str(chart.series[0].val.numRef.f)
                except Exception:
                    pass
                print(f"FAIL: Component 5 — Chart series not extended to row 10, found: {ref_info}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
