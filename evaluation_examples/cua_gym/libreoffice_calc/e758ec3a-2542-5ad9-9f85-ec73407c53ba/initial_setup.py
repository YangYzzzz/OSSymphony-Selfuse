"""
Initial Setup: Create a ProductSales sheet with 360 rows of sales data
Task ID: calc_pivot_031
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_031'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ProductSales'

    # Headers
    headers = ['SaleID', 'SaleDate', 'ProductLine', 'Units', 'Revenue']
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align

    # Monthly revenue targets per product line (sum = 385000)
    # Ground truth: Jan/Standard=4200, Jan/Premium=8500, Jan/Enterprise=15000
    monthly_targets = {
        1:  {'Standard': 4200,  'Premium': 8500,  'Enterprise': 15000},
        2:  {'Standard': 3500,  'Premium': 9800,  'Enterprise': 18700},
        3:  {'Standard': 3200,  'Premium': 10200, 'Enterprise': 19500},
        4:  {'Standard': 3800,  'Premium': 9500,  'Enterprise': 18200},
        5:  {'Standard': 3600,  'Premium': 10800, 'Enterprise': 19800},
        6:  {'Standard': 3400,  'Premium': 9200,  'Enterprise': 17900},
        7:  {'Standard': 3700,  'Premium': 10500, 'Enterprise': 19200},
        8:  {'Standard': 3300,  'Premium': 9900,  'Enterprise': 18500},
        9:  {'Standard': 3900,  'Premium': 10100, 'Enterprise': 19000},
        10: {'Standard': 3500,  'Premium': 9700,  'Enterprise': 18800},
        11: {'Standard': 3600,  'Premium': 10300, 'Enterprise': 19600},
        12: {'Standard': 4100,  'Premium': 11000, 'Enterprise': 17500},
    }

    product_lines = ['Standard', 'Premium', 'Enterprise']
    rows = []
    sale_id = 1

    days_in_month = {
        1: 31, 2: 29, 3: 31, 4: 30, 5: 31, 6: 30,
        7: 31, 8: 31, 9: 30, 10: 31, 11: 30, 12: 31
    }

    for month in range(1, 13):
        dim = days_in_month[month]
        for pl in product_lines:
            target_rev = monthly_targets[month][pl]
            num_sales = 10
            base_rev = target_rev // num_sales
            remainder = target_rev - base_rev * num_sales

            for i in range(num_sales):
                day = random.randint(1, dim)
                sale_date = date(2024, month, day)
                rev = base_rev + (1 if i < remainder else 0)
                if pl == 'Standard':
                    unit_price = random.choice([35, 40, 42, 45, 50])
                elif pl == 'Premium':
                    unit_price = random.choice([75, 80, 85, 90, 95])
                else:
                    unit_price = random.choice([150, 160, 175, 180, 200])
                units = max(1, round(rev / unit_price))
                rows.append((sale_id, sale_date, pl, units, rev))
                sale_id += 1

    # Shuffle for realism, then reassign sequential IDs
    random.shuffle(rows)
    for i in range(len(rows)):
        _, sd, pl, u, rev = rows[i]
        rows[i] = (i + 1, sd, pl, u, rev)

    # Write data
    for r, (sid, sale_date, pl, units, rev) in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=sid)
        ws.cell(row=r, column=2, value=sale_date)
        ws.cell(row=r, column=2).number_format = 'MM/DD/YYYY'
        ws.cell(row=r, column=3, value=pl)
        ws.cell(row=r, column=4, value=units)
        ws.cell(row=r, column=5, value=rev)
        ws.cell(row=r, column=5).number_format = '#,##0'

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 14

    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    verify_total = sum(row[4] for row in rows)
    print(f'Initial file created: {OUTPUT}')
    print(f'Total revenue: {verify_total} (expected: 385000)')
    print(f'Total rows: {len(rows)} (expected: 360)')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
