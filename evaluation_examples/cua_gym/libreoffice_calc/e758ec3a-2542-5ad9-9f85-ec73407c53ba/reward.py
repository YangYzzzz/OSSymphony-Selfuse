"""
Reward Script: Create a pivot table showing monthly revenue trends for each product line
Task ID: calc_pivot_031
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20): PivotTable sheet exists (not in initial)
  Component 2 (0.30): Correct structure — Month rows, ProductLine columns, proper headers
  Component 3 (0.30): Key data values match expected ground truth
  Component 4 (0.20): Grand Total row is correct with annual total 385000
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_031'

# Expected monthly data: {month_name: {Standard, Premium, Enterprise, Grand Total}}
EXPECTED_DATA = {
    'January':   {'Standard': 4200,  'Premium': 8500,  'Enterprise': 15000, 'Grand Total': 27700},
    'February':  {'Standard': 3500,  'Premium': 9800,  'Enterprise': 18700, 'Grand Total': 32000},
    'March':     {'Standard': 3200,  'Premium': 10200, 'Enterprise': 19500, 'Grand Total': 32900},
    'April':     {'Standard': 3800,  'Premium': 9500,  'Enterprise': 18200, 'Grand Total': 31500},
    'May':       {'Standard': 3600,  'Premium': 10800, 'Enterprise': 19800, 'Grand Total': 34200},
    'June':      {'Standard': 3400,  'Premium': 9200,  'Enterprise': 17900, 'Grand Total': 30500},
    'July':      {'Standard': 3700,  'Premium': 10500, 'Enterprise': 19200, 'Grand Total': 33400},
    'August':    {'Standard': 3300,  'Premium': 9900,  'Enterprise': 18500, 'Grand Total': 31700},
    'September': {'Standard': 3900,  'Premium': 10100, 'Enterprise': 19000, 'Grand Total': 33000},
    'October':   {'Standard': 3500,  'Premium': 9700,  'Enterprise': 18800, 'Grand Total': 32000},
    'November':  {'Standard': 3600,  'Premium': 10300, 'Enterprise': 19600, 'Grand Total': 33500},
    'December':  {'Standard': 4100,  'Premium': 11000, 'Enterprise': 17500, 'Grand Total': 32600},
}

EXPECTED_GRAND_TOTALS = {'Standard': 43800, 'Premium': 119500, 'Enterprise': 221700, 'Grand Total': 385000}

MONTH_NAMES = [
    'January', 'February', 'March', 'April', 'May', 'June',
    'July', 'August', 'September', 'October', 'November', 'December'
]


def find_pivot_sheet(wb):
    """Find a sheet that looks like a pivot table (not ProductSales)."""
    for sn in wb.sheetnames:
        if sn.lower() != 'productsales':
            return wb[sn]
    return None


def find_header_row(ws):
    """Find the row containing 'Month' or similar header, and map column positions."""
    for r in range(1, min(ws.max_row + 1, 10)):
        for c in range(1, min(ws.max_column + 1, 10)):
            val = ws.cell(r, c).value
            if val and str(val).strip().lower() == 'month':
                # Found header row, map columns
                col_map = {}
                for cc in range(1, ws.max_column + 1):
                    hdr = ws.cell(r, cc).value
                    if hdr:
                        col_map[str(hdr).strip()] = cc
                return r, col_map
    return None, None


def _find_value_in_row(ws, row, col_map, target, tolerance=1):
    """Check if target value exists in the given row, first in Grand Total col, then any col."""
    if 'Grand Total' in col_map:
        gt_col = col_map['Grand Total']
        gt_val = ws.cell(row, gt_col).value
        if gt_val is not None:
            try:
                if abs(float(gt_val) - target) < tolerance:
                    return float(gt_val)
            except (ValueError, TypeError):
                pass
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row, c).value
        if val is not None:
            try:
                if abs(float(val) - target) < tolerance:
                    return float(val)
            except (ValueError, TypeError):
                pass
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: PivotTable sheet exists (0.2 points)
    # This FAILS on initial (only has ProductSales) and PASSES on golden
    try:
        pivot_ws = find_pivot_sheet(wb)
        if pivot_ws is not None:
            print(f"PASS: Component 1 — Pivot table sheet found: '{pivot_ws.title}' (0.2 pts)")
            total_score += 0.2
        else:
            print("FAIL: Component 1 — No pivot table sheet found (only ProductSales exists)")
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Component 2: Correct structure — headers with Month, Standard, Premium, Enterprise (0.3 points)
    try:
        header_row, col_map = find_header_row(pivot_ws)
        if header_row is None:
            print("FAIL: Component 2 — No 'Month' header found in pivot sheet")
        else:
            required_cols = ['Standard', 'Premium', 'Enterprise']
            found_cols = [c for c in required_cols if c in col_map]
            has_month = 'Month' in col_map

            if has_month and len(found_cols) == len(required_cols):
                # Check that 12 month rows exist
                month_col = col_map['Month']
                found_months = []
                for r in range(header_row + 1, pivot_ws.max_row + 1):
                    val = pivot_ws.cell(r, month_col).value
                    if val and str(val).strip() in MONTH_NAMES:
                        found_months.append(str(val).strip())

                if len(found_months) == 12:
                    print(f"PASS: Component 2 — Correct structure with Month rows and ProductLine columns, 12 months found (0.3 pts)")
                    total_score += 0.3
                elif len(found_months) >= 6:
                    partial = 0.15
                    print(f"PARTIAL: Component 2 — Found {len(found_months)}/12 months ({partial} pts)")
                    total_score += partial
                else:
                    print(f"FAIL: Component 2 — Only {len(found_months)}/12 months found")
            else:
                missing = [c for c in required_cols if c not in col_map]
                print(f"FAIL: Component 2 — Missing columns: {missing}, has Month: {has_month}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Key data values match expected ground truth (0.3 points)
    # Check a selection of cells: Jan/Standard=4200, Jan/Premium=8500, Jan/Enterprise=15000
    # and a sampling across other months
    try:
        header_row, col_map = find_header_row(pivot_ws)
        if header_row is None or 'Month' not in col_map:
            print("FAIL: Component 3 — Cannot verify values without proper headers")
        else:
            month_col = col_map['Month']
            # Build month-to-row mapping
            month_row_map = {}
            for r in range(header_row + 1, pivot_ws.max_row + 1):
                val = pivot_ws.cell(r, month_col).value
                if val and str(val).strip() in MONTH_NAMES:
                    month_row_map[str(val).strip()] = r

            # Check key values from context: Jan/Standard=4200, Jan/Premium=8500, Jan/Enterprise=15000
            # Plus spot checks across other months for robustness
            checks_to_verify = [
                ('January', 'Standard', 4200),
                ('January', 'Premium', 8500),
                ('January', 'Enterprise', 15000),
                ('June', 'Standard', 3400),
                ('June', 'Enterprise', 17900),
                ('December', 'Premium', 11000),
            ]

            passed = 0
            total_checks = len(checks_to_verify)
            for month, product, expected_val in checks_to_verify:
                if month not in month_row_map:
                    print(f"  FAIL: {month}/{product} — month row not found")
                    continue
                if product not in col_map:
                    print(f"  FAIL: {month}/{product} — column not found")
                    continue
                r = month_row_map[month]
                c = col_map[product]
                actual = pivot_ws.cell(r, c).value
                try:
                    if actual is not None and abs(float(actual) - expected_val) < 1:
                        passed += 1
                    else:
                        print(f"  FAIL: {month}/{product} — expected {expected_val}, got {actual}")
                except (ValueError, TypeError):
                    print(f"  FAIL: {month}/{product} — expected {expected_val}, got {actual} (type error)")

            if passed == total_checks:
                print(f"PASS: Component 3 — All {total_checks} key values correct (0.3 pts)")
                total_score += 0.3
            elif passed >= total_checks // 2:
                partial = round(0.3 * passed / total_checks, 2)
                print(f"PARTIAL: Component 3 — {passed}/{total_checks} values correct ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 3 — Only {passed}/{total_checks} values correct")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Grand Total row correct with annual total = 385000 (0.2 points)
    try:
        header_row, col_map = find_header_row(pivot_ws)
        if header_row is None:
            print("FAIL: Component 4 — No headers found")
        else:
            month_col = col_map['Month']
            grand_total_row = None
            for r in range(header_row + 1, pivot_ws.max_row + 1):
                val = pivot_ws.cell(r, month_col).value
                if val and 'total' in str(val).strip().lower():
                    grand_total_row = r
                    break

            if grand_total_row is None:
                # Also check if grand total is in any other column position
                for r in range(header_row + 1, pivot_ws.max_row + 1):
                    for c in range(1, pivot_ws.max_column + 1):
                        val = pivot_ws.cell(r, c).value
                        if val and 'grand total' in str(val).strip().lower():
                            grand_total_row = r
                            break
                    if grand_total_row:
                        break

            if grand_total_row is None:
                print("FAIL: Component 4 — No Grand Total row found")
            else:
                # Find the grand total value (should be 385000)
                gt_value_found = _find_value_in_row(pivot_ws, grand_total_row, col_map, 385000)
                if gt_value_found:
                    print(f"PASS: Component 4 — Grand Total = 385000 (0.2 pts)")
                    total_score += 0.2
                else:
                    print(f"FAIL: Component 4 — Grand Total of 385000 not found in row {grand_total_row}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
