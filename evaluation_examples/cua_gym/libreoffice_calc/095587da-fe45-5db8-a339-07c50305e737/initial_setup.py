"""
Initial Setup: Create data_pipeline.xlsx with RawData (200 rows) and empty Log sheet,
plus /tmp/new_data.csv with 50 rows (some overlapping IDs).
Task ID: calc_gg5_033
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import csv
import random
from datetime import datetime, timedelta

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_033'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'
CSV_PATH = '/tmp/new_data.csv'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

# Realistic data pools
SOURCES = ['API', 'WebScraper', 'ManualEntry', 'BatchUpload', 'SensorFeed', 'ETL_Pipeline']
STATUSES = ['Validated', 'Pending', 'Flagged', 'Processed', 'Review']
DEPARTMENTS = ['Engineering', 'Sales', 'Marketing', 'Finance', 'Operations', 'Support', 'HR']

random.seed(42)

def random_timestamp(start_year=2023, end_year=2024):
    start = datetime(start_year, 1, 1)
    end = datetime(end_year, 12, 31)
    delta = end - start
    rand_days = random.randint(0, delta.days)
    rand_secs = random.randint(0, 86399)
    dt = start + timedelta(days=rand_days, seconds=rand_secs)
    return dt.strftime('%Y-%m-%d %H:%M:%S')

def random_value():
    return round(random.uniform(10.0, 9999.99), 2)

def create_initial():
    wb = openpyxl.Workbook()

    # --- RawData sheet ---
    ws = wb.active
    ws.title = 'RawData'

    headers = ['UniqueID', 'Source', 'Timestamp', 'Value', 'Status']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 200 rows of data (rows 2-201), IDs: REC-0001 to REC-0200
    existing_ids = []
    for i in range(1, 201):
        row = i + 1
        uid = f'REC-{i:04d}'
        existing_ids.append(uid)
        ws.cell(row=row, column=1, value=uid)
        ws.cell(row=row, column=2, value=random.choice(SOURCES))
        ws.cell(row=row, column=3, value=random_timestamp())
        ws.cell(row=row, column=4, value=random_value())
        ws.cell(row=row, column=5, value=random.choice(STATUSES))

    # --- Log sheet (empty, A1 must be empty) ---
    ws_log = wb.create_sheet('Log')
    # Intentionally leave A1 empty

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # --- Create /tmp/new_data.csv with 50 rows ---
    # 15 overlapping IDs (from existing REC-0050 to REC-0064) and 35 new IDs (REC-0201 to REC-0235)
    csv_rows = []
    # Overlapping rows (IDs that already exist)
    for i in range(50, 65):
        uid = f'REC-{i:04d}'
        csv_rows.append([uid, random.choice(SOURCES), random_timestamp(2024, 2024),
                         random_value(), random.choice(STATUSES)])
    # New rows
    for i in range(201, 236):
        uid = f'REC-{i:04d}'
        csv_rows.append([uid, random.choice(SOURCES), random_timestamp(2024, 2024),
                         random_value(), random.choice(STATUSES)])

    random.shuffle(csv_rows)

    with open(CSV_PATH, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)  # Same headers as RawData
        writer.writerows(csv_rows)

    print(f'CSV file created: {CSV_PATH} with {len(csv_rows)} data rows')

    # GUI-ready: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
