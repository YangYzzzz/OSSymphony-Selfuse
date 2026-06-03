"""
Reward Script: CSV Import with Deduplication and Log Entry
Task ID: calc_gg5_033
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Row count increased from 200 to >200 (new data appended)
  Component 2 (0.25): No duplicate UniqueIDs in column A (deduplication)
  Component 3 (0.25): Log sheet A1 contains import completion timestamp message
  Component 4 (0.20): Original 200 rows preserved AND new data appended (compound check)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_033'


def persist_app_state(domain):
    """Save any unsaved LibreOffice changes before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: RawData sheet must exist
    if 'RawData' not in wb.sheetnames:
        print("CRITICAL: 'RawData' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Log sheet must exist
    if 'Log' not in wb.sheetnames:
        print("CRITICAL: 'Log' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws_raw = wb['RawData']
    ws_log = wb['Log']

    # Component 1: Row count increased — new data was appended (0.30 points)
    # Initial has 201 rows (1 header + 200 data). After import, should be >201.
    # Golden has 236 rows (200 original + 35 new non-duplicate rows).
    try:
        max_row = ws_raw.max_row
        data_rows = max_row - 1  # subtract header
        print(f"INFO: RawData max_row={max_row}, data_rows={data_rows}")

        if data_rows > 200:
            new_rows = data_rows - 200
            print(f"PASS: Component 1 — {new_rows} new rows appended (total {data_rows} data rows) (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 — Expected >200 data rows, found {data_rows}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: No duplicate UniqueIDs in column A (0.25 points)
    # After deduplication, all values in column A (rows 2+) should be unique.
    try:
        ids = []
        for r in range(2, ws_raw.max_row + 1):
            val = ws_raw.cell(row=r, column=1).value
            if val is not None:
                ids.append(str(val).strip())

        unique_ids = set(ids)
        if len(ids) > 200 and len(ids) == len(unique_ids):
            print(f"PASS: Component 2 — All {len(ids)} IDs are unique, no duplicates (0.25 pts)")
            total_score += 0.25
        elif len(ids) <= 200:
            print(f"FAIL: Component 2 — Only {len(ids)} IDs found, no new data appended yet")
        else:
            dup_count = len(ids) - len(unique_ids)
            print(f"FAIL: Component 2 — Found {dup_count} duplicate IDs among {len(ids)} rows")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Log sheet A1 contains timestamp import message (0.25 points)
    # Expected format: "YYYY-MM-DD HH:MM:SS - Import complete: XX rows added"
    try:
        log_val = ws_log['A1'].value
        if log_val is not None:
            log_str = str(log_val).strip()
            # Check that it contains key parts of the expected log message
            has_timestamp = False
            has_import_msg = False

            # Check for date-like pattern (YYYY-MM-DD or similar)
            import re
            if re.search(r'\d{4}-\d{2}-\d{2}', log_str):
                has_timestamp = True

            # Check for import completion message
            lower_log = log_str.lower()
            if 'import' in lower_log and ('complete' in lower_log or 'done' in lower_log or 'finish' in lower_log):
                has_import_msg = True

            if has_timestamp and has_import_msg:
                print(f"PASS: Component 3 — Log A1 contains timestamp + import message: '{log_str}' (0.25 pts)")
                total_score += 0.25
            elif has_import_msg:
                print(f"PARTIAL: Component 3 — Log has import message but no timestamp: '{log_str}' (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 3 — Log A1 does not match expected pattern: '{log_str}'")
        else:
            print("FAIL: Component 3 — Log A1 is empty (None)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: New data appended AND original data preserved (0.20 points)
    # This is a compound check: new rows must exist AND original rows must be intact.
    # Row 2 should still be REC-0001, row 201 should still be REC-0200,
    # AND there must be data beyond row 201 (new rows appended).
    try:
        first_id = ws_raw.cell(row=2, column=1).value
        last_orig_id = ws_raw.cell(row=201, column=1).value
        header_a = ws_raw.cell(row=1, column=1).value

        first_ok = (str(first_id).strip() == 'REC-0001') if first_id else False
        last_ok = (str(last_orig_id).strip() == 'REC-0200') if last_orig_id else False
        header_ok = (str(header_a).strip() == 'UniqueID') if header_a else False
        has_new_data = ws_raw.max_row > 201

        if first_ok and last_ok and header_ok and has_new_data:
            print(f"PASS: Component 4 — Original data intact + new data present: header='UniqueID', row2='{first_id}', row201='{last_orig_id}', max_row={ws_raw.max_row} (0.20 pts)")
            total_score += 0.20
        elif not has_new_data:
            print(f"FAIL: Component 4 — No new data appended (max_row={ws_raw.max_row}, need >201)")
        else:
            details = f"header={header_a}, row2={first_id}, row201={last_orig_id}"
            print(f"FAIL: Component 4 — Original data corrupted: {details}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist app state before verification
persist_app_state("libreoffice_calc")

# Run verification
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
