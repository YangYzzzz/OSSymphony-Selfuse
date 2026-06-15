"""
Reward Script: Vendor Invoice Reconciliation Spreadsheet
Task ID: calc_grs_036
Domain: libreoffice_calc
Scoring:
  C1 (0.25) - Reconciliation sheet has VLOOKUP formulas linking invoices to POs
  C2 (0.20) - Quantity and price discrepancy columns calculated correctly
  C3 (0.15) - Invoice Total and Total Discrepancy columns (K, L) present with formulas
  C4 (0.15) - Summary section with Total PO Value, Total Invoiced, Total Discrepancy, % with discrepancies
  C5 (0.15) - Conditional formatting highlights rows with discrepancies in red
  C6 (0.10) - Currency columns formatted consistently ($#,##0.00)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_036'


def persist_app_state(domain: str):
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify vendor invoice reconciliation spreadsheet.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Must have a Reconciliation sheet
    if 'Reconciliation' not in wb.sheetnames:
        # Check for alternative names
        recon_sheet = None
        for name in wb.sheetnames:
            if 'reconcil' in name.lower() or 'recon' in name.lower():
                recon_sheet = name
                break
        if recon_sheet is None:
            print("FAIL: No Reconciliation sheet found")
            print("REWARD: 0.0")
            return 0.0
        ws = wb[recon_sheet]
        print(f"INFO: Using sheet '{recon_sheet}' as reconciliation sheet")
    else:
        ws = wb['Reconciliation']

    # Also need Purchase Orders and Vendor Invoices as preconditions
    if 'Purchase Orders' not in wb.sheetnames or 'Vendor Invoices' not in wb.sheetnames:
        print("FAIL: Missing Purchase Orders or Vendor Invoices sheet")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: VLOOKUP formulas in Reconciliation linking invoices to POs (0.25 pts)
    # The golden has VLOOKUP in D2:D13 fetching PO Quantity and G2:G13 fetching PO Unit Price
    try:
        vlookup_count = 0
        expected_vlookup_cells = 0
        for row in range(2, 14):  # rows 2-13 (12 invoice rows)
            for col_letter in ['D', 'G', 'J']:  # PO Quantity, PO Unit Price, PO Total via VLOOKUP
                expected_vlookup_cells += 1
                cell = ws[f'{col_letter}{row}']
                val = cell.value
                if val and isinstance(val, str) and 'VLOOKUP' in val.upper():
                    vlookup_count += 1

        if vlookup_count >= 20:  # At least 20 of 36 expected VLOOKUP cells
            print(f"PASS: Component 1 - Found {vlookup_count} VLOOKUP formulas in reconciliation (0.25 pts)")
            total_score += 0.25
        elif vlookup_count >= 12:
            partial = 0.15
            print(f"PARTIAL: Component 1 - Found {vlookup_count} VLOOKUP formulas (partial: {partial} pts)")
            total_score += partial
        elif vlookup_count >= 6:
            partial = 0.08
            print(f"PARTIAL: Component 1 - Found {vlookup_count} VLOOKUP formulas (partial: {partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - Only {vlookup_count} VLOOKUP formulas found (need >= 20)")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Quantity and Price discrepancy columns (0.20 pts)
    # F column should compute quantity discrepancy (Invoice Qty - PO Qty)
    # I column should compute price discrepancy (Invoice Price - PO Price)
    try:
        disc_formula_count = 0
        for row in range(2, 14):
            # Check F column (quantity discrepancy)
            f_val = ws[f'F{row}'].value
            if f_val and isinstance(f_val, str) and ('E' in f_val.upper() and 'D' in f_val.upper() and '-' in f_val):
                disc_formula_count += 1
            elif f_val and isinstance(f_val, str) and any(op in f_val for op in ['-', '+', 'IF']):
                disc_formula_count += 1

            # Check I column (price discrepancy)
            i_val = ws[f'I{row}'].value
            if i_val and isinstance(i_val, str) and ('H' in i_val.upper() and 'G' in i_val.upper() and '-' in i_val):
                disc_formula_count += 1
            elif i_val and isinstance(i_val, str) and any(op in i_val for op in ['-', '+', 'IF']):
                disc_formula_count += 1

        if disc_formula_count >= 18:  # 24 expected (12 rows x 2 columns)
            print(f"PASS: Component 2 - {disc_formula_count} discrepancy formulas found (0.20 pts)")
            total_score += 0.20
        elif disc_formula_count >= 10:
            partial = 0.10
            print(f"PARTIAL: Component 2 - {disc_formula_count} discrepancy formulas (partial: {partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 - Only {disc_formula_count} discrepancy formulas found (need >= 18)")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Invoice Total (K) and Total Discrepancy (L) columns with formulas (0.15 pts)
    try:
        kl_formula_count = 0
        for row in range(2, 14):
            k_val = ws.cell(row=row, column=11).value  # K column
            l_val = ws.cell(row=row, column=12).value  # L column
            if k_val and isinstance(k_val, str) and '=' in str(k_val):
                kl_formula_count += 1
            if l_val and isinstance(l_val, str) and '=' in str(l_val):
                kl_formula_count += 1

        if kl_formula_count >= 18:  # 24 expected
            print(f"PASS: Component 3 - {kl_formula_count} K/L formulas found (0.15 pts)")
            total_score += 0.15
        elif kl_formula_count >= 10:
            partial = 0.08
            print(f"PARTIAL: Component 3 - {kl_formula_count} K/L formulas (partial: {partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 - Only {kl_formula_count} K/L column formulas found (need >= 18)")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Summary section with aggregate formulas (0.15 pts)
    # Expect: Total PO Value (SUM), Total Invoiced Amount (SUM),
    #         Total Discrepancy Value (SUM), Percentage with discrepancies (COUNTIF)
    try:
        summary_items_found = 0

        # Scan rows 14-30 for summary labels and formulas
        found_labels = {}
        for row in range(14, min(ws.max_row + 1, 35)):
            label = ws.cell(row=row, column=1).value
            formula = ws.cell(row=row, column=2).value
            if label and isinstance(label, str):
                label_lower = label.lower()
                if formula and isinstance(formula, str):
                    formula_upper = formula.upper()
                    if 'total po' in label_lower or 'po value' in label_lower:
                        if 'SUM' in formula_upper:
                            summary_items_found += 1
                            found_labels['total_po'] = True
                    if 'invoiced' in label_lower or 'invoice total' in label_lower or 'invoice amount' in label_lower:
                        if 'SUM' in formula_upper:
                            summary_items_found += 1
                            found_labels['total_inv'] = True
                    if 'discrepancy' in label_lower and ('total' in label_lower or 'value' in label_lower):
                        if 'SUM' in formula_upper:
                            summary_items_found += 1
                            found_labels['total_disc'] = True
                    if 'percent' in label_lower or '%' in label_lower:
                        if 'COUNTIF' in formula_upper or 'COUNT' in formula_upper:
                            summary_items_found += 1
                            found_labels['pct_disc'] = True

        if summary_items_found >= 4:
            print(f"PASS: Component 4 - All 4 summary items found (0.15 pts)")
            total_score += 0.15
        elif summary_items_found >= 2:
            partial = round(0.15 * summary_items_found / 4, 2)
            print(f"PARTIAL: Component 4 - {summary_items_found}/4 summary items (partial: {partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 - Only {summary_items_found}/4 summary items found. Labels checked: {found_labels}")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Conditional formatting highlighting discrepancy rows in red (0.15 pts)
    try:
        has_cond_fmt = False
        red_rule_found = False

        cf_rules = ws.conditional_formatting
        rule_count = 0
        for cf in cf_rules:
            rule_count += 1
            for rule in cf.rules:
                # Check if it involves discrepancy columns and uses red fill
                if rule.dxf and rule.dxf.fill:
                    fill_color = None
                    try:
                        fill_color = rule.dxf.fill.fgColor.rgb
                    except:
                        pass
                    # Check if the color is reddish (FFFF0000 or similar)
                    if fill_color and isinstance(fill_color, str):
                        # Red = high R, low G, low B
                        if fill_color.upper().startswith('FFFF') and fill_color.upper()[4:6] == '00':
                            red_rule_found = True
                        elif 'FF0000' in fill_color.upper():
                            red_rule_found = True
                        elif fill_color.upper() in ['FFFF0000']:
                            red_rule_found = True

                # Also check if formula references discrepancy columns
                if rule.formula:
                    formula_str = str(rule.formula)
                    if any(ref in formula_str.upper() for ref in ['$F', '$I', '$L', 'F2', 'I2', 'L2']):
                        has_cond_fmt = True

        if red_rule_found and (has_cond_fmt or rule_count > 0):
            print(f"PASS: Component 5 - Conditional formatting with red fill for discrepancies (0.15 pts)")
            total_score += 0.15
        elif rule_count > 0:
            # Some conditional formatting exists but may not be exactly right
            partial = 0.08
            print(f"PARTIAL: Component 5 - Conditional formatting exists but red fill check uncertain ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 - No conditional formatting rules found on Reconciliation sheet")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    # Component 6: Currency columns formatted consistently with $#,##0.00 (0.10 pts)
    # In golden, PO E/F, Invoice E/F, and Recon G/H/I/J/K/B17-B19 all use $#,##0.00
    # In initial, these are all General format
    try:
        currency_count = 0
        total_checked = 0

        # Check Purchase Orders sheet E and F columns
        ws_po = wb['Purchase Orders']
        for row in range(2, 14):
            for col_letter in ['E', 'F']:
                total_checked += 1
                fmt = ws_po[f'{col_letter}{row}'].number_format
                if '$' in str(fmt):
                    currency_count += 1

        # Check Vendor Invoices sheet E and F columns
        ws_inv = wb['Vendor Invoices']
        for row in range(2, 14):
            for col_letter in ['E', 'F']:
                total_checked += 1
                fmt = ws_inv[f'{col_letter}{row}'].number_format
                if '$' in str(fmt):
                    currency_count += 1

        # Check Reconciliation currency columns
        for row in range(2, 14):
            for col_letter in ['G', 'J', 'K']:
                total_checked += 1
                fmt = ws[f'{col_letter}{row}'].number_format
                if '$' in str(fmt):
                    currency_count += 1

        pct = currency_count / max(total_checked, 1)
        if pct >= 0.6:
            print(f"PASS: Component 6 - {currency_count}/{total_checked} cells have currency format (0.10 pts)")
            total_score += 0.10
        elif pct >= 0.3:
            partial = 0.05
            print(f"PARTIAL: Component 6 - {currency_count}/{total_checked} cells have currency format ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 6 - Only {currency_count}/{total_checked} cells have currency format")
    except Exception as e:
        print(f"ERROR: Component 6 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
