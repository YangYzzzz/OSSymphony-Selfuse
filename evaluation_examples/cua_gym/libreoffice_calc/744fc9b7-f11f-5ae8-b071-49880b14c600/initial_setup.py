"""
Initial Setup: Vendor Invoice Reconciliation Spreadsheet
Task ID: calc_grs_036
Domain: libreoffice_calc

Creates Sheet1 (Purchase Orders), Sheet2 (Vendor Invoices), and a blank Sheet3
(Reconciliation) that the agent must populate with VLOOKUPs, conditional formatting,
and a summary section.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_036'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Purchase Orders ---
    ws1 = wb.active
    ws1.title = 'Purchase Orders'

    headers1 = ['PO Number', 'Vendor', 'Item Description', 'Ordered Quantity', 'Unit Price', 'PO Total']
    header_font = Font(bold=True)
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font

    po_data = [
        ['PO-2025-001', 'Apex Industrial Supply', 'Stainless Steel Bearings (Box of 50)', 120, 34.75, 4170.00],
        ['PO-2025-002', 'GlobalTech Components', 'Circuit Board Assembly PCB-440', 80, 127.50, 10200.00],
        ['PO-2025-003', 'Meridian Office Solutions', 'Ergonomic Desk Chair Model K9', 25, 489.00, 12225.00],
        ['PO-2025-004', 'Apex Industrial Supply', 'Hydraulic Cylinder HC-200', 40, 215.00, 8600.00],
        ['PO-2025-005', 'Pinnacle Chemical Corp', 'Industrial Solvent Grade A (5L)', 200, 18.50, 3700.00],
        ['PO-2025-006', 'GlobalTech Components', 'LED Display Panel 15-inch', 60, 312.00, 18720.00],
        ['PO-2025-007', 'Summit Logistics Partners', 'Pallet Wrap Heavy Duty (Roll)', 500, 8.25, 4125.00],
        ['PO-2025-008', 'Meridian Office Solutions', 'Wireless Keyboard and Mouse Set', 150, 45.99, 6898.50],
        ['PO-2025-009', 'Pinnacle Chemical Corp', 'Epoxy Resin Kit ER-100', 75, 62.00, 4650.00],
        ['PO-2025-010', 'Apex Industrial Supply', 'Tungsten Carbide Drill Bits (Set)', 30, 178.50, 5355.00],
        ['PO-2025-011', 'GlobalTech Components', 'Fiber Optic Cable Cat6a (100m)', 45, 89.95, 4047.75],
        ['PO-2025-012', 'Summit Logistics Partners', 'Corrugated Shipping Boxes (Bundle)', 300, 3.50, 1050.00],
    ]

    for r, row_data in enumerate(po_data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws1.column_dimensions['A'].width = 16
    ws1.column_dimensions['B'].width = 28
    ws1.column_dimensions['C'].width = 38
    ws1.column_dimensions['D'].width = 18
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 12

    # --- Sheet 2: Vendor Invoices ---
    ws2 = wb.create_sheet('Vendor Invoices')

    headers2 = ['Invoice Number', 'PO Number', 'Vendor', 'Invoiced Quantity', 'Unit Price', 'Invoice Total']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font

    # Some invoices match POs exactly, some have discrepancies in quantity or price
    invoice_data = [
        ['INV-8801', 'PO-2025-001', 'Apex Industrial Supply', 120, 34.75, 4170.00],        # exact match
        ['INV-8802', 'PO-2025-002', 'GlobalTech Components', 85, 127.50, 10837.50],         # qty discrepancy: 85 vs 80
        ['INV-8803', 'PO-2025-003', 'Meridian Office Solutions', 25, 499.00, 12475.00],     # price discrepancy: 499 vs 489
        ['INV-8804', 'PO-2025-004', 'Apex Industrial Supply', 40, 215.00, 8600.00],         # exact match
        ['INV-8805', 'PO-2025-005', 'Pinnacle Chemical Corp', 210, 19.25, 4042.50],         # qty + price discrepancy
        ['INV-8806', 'PO-2025-006', 'GlobalTech Components', 60, 312.00, 18720.00],         # exact match
        ['INV-8807', 'PO-2025-007', 'Summit Logistics Partners', 500, 8.75, 4375.00],       # price discrepancy: 8.75 vs 8.25
        ['INV-8808', 'PO-2025-008', 'Meridian Office Solutions', 145, 45.99, 6668.55],      # qty discrepancy: 145 vs 150
        ['INV-8809', 'PO-2025-009', 'Pinnacle Chemical Corp', 75, 62.00, 4650.00],          # exact match
        ['INV-8810', 'PO-2025-010', 'Apex Industrial Supply', 35, 185.00, 6475.00],         # qty + price discrepancy
        ['INV-8811', 'PO-2025-011', 'GlobalTech Components', 45, 89.95, 4047.75],           # exact match
        ['INV-8812', 'PO-2025-012', 'Summit Logistics Partners', 300, 3.50, 1050.00],       # exact match
    ]

    for r, row_data in enumerate(invoice_data, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 28
    ws2.column_dimensions['D'].width = 18
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 14

    # --- Sheet 3: Reconciliation (blank - agent must build this) ---
    ws3 = wb.create_sheet('Reconciliation')
    # Leave blank - the task requires the agent to populate this sheet

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
