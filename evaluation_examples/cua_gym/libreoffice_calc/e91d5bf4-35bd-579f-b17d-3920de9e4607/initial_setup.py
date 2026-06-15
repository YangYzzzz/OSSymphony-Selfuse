"""
Initial Setup: Workbook with structure protection enabled
Task ID: calc_ps_016
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_016'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Common styles ---
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # --- Sheet 1: Data ---
    ws_data = wb.active
    ws_data.title = 'Data'

    data_headers = ['Employee ID', 'Name', 'Department', 'Position', 'Salary', 'Start Date', 'Performance Rating']
    for col, h in enumerate(data_headers, 1):
        cell = ws_data.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    employee_data = [
        ['EMP001', 'Sarah Chen', 'Engineering', 'Senior Developer', 95000, '2021-03-15', 4.5],
        ['EMP002', 'Marcus Johnson', 'Marketing', 'Campaign Manager', 72000, '2022-06-01', 3.8],
        ['EMP003', 'Aisha Patel', 'Engineering', 'Tech Lead', 112000, '2019-11-20', 4.8],
        ['EMP004', 'James O\'Brien', 'Finance', 'Financial Analyst', 68000, '2023-01-10', 3.5],
        ['EMP005', 'Maria Garcia', 'Human Resources', 'HR Specialist', 65000, '2022-09-05', 4.0],
        ['EMP006', 'David Kim', 'Engineering', 'Junior Developer', 62000, '2024-02-18', 3.2],
        ['EMP007', 'Elena Volkov', 'Marketing', 'Content Director', 88000, '2020-07-12', 4.3],
        ['EMP008', 'Robert Taylor', 'Finance', 'Controller', 105000, '2018-04-22', 4.6],
        ['EMP009', 'Priya Sharma', 'Engineering', 'QA Engineer', 78000, '2021-08-30', 3.9],
        ['EMP010', 'Thomas Mueller', 'Operations', 'Logistics Manager', 82000, '2020-01-07', 4.1],
        ['EMP011', 'Lisa Nakamura', 'Engineering', 'DevOps Engineer', 91000, '2022-03-14', 4.4],
        ['EMP012', 'Carlos Rivera', 'Marketing', 'Brand Strategist', 76000, '2023-05-20', 3.7],
    ]

    for r, row_data in enumerate(employee_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_data.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 5:  # Salary column
                cell.number_format = '$#,##0'
            elif c == 7:  # Rating column
                cell.number_format = '0.0'

    # Set column widths
    ws_data.column_dimensions['A'].width = 12
    ws_data.column_dimensions['B'].width = 20
    ws_data.column_dimensions['C'].width = 18
    ws_data.column_dimensions['D'].width = 20
    ws_data.column_dimensions['E'].width = 12
    ws_data.column_dimensions['F'].width = 14
    ws_data.column_dimensions['G'].width = 18

    # --- Sheet 2: Charts ---
    ws_charts = wb.create_sheet('Charts')

    chart_headers = ['Department', 'Headcount', 'Avg Salary', 'Budget Allocation']
    for col, h in enumerate(chart_headers, 1):
        cell = ws_charts.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    dept_data = [
        ['Engineering', 5, 87600, 520000],
        ['Marketing', 3, 78667, 280000],
        ['Finance', 2, 86500, 195000],
        ['Human Resources', 1, 65000, 85000],
        ['Operations', 1, 82000, 110000],
    ]

    for r, row_data in enumerate(dept_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_charts.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c in (3, 4):
                cell.number_format = '$#,##0'

    ws_charts.column_dimensions['A'].width = 18
    ws_charts.column_dimensions['B'].width = 12
    ws_charts.column_dimensions['C'].width = 14
    ws_charts.column_dimensions['D'].width = 18

    # --- Sheet 3: Summary ---
    ws_summary = wb.create_sheet('Summary')

    summary_headers = ['Metric', 'Value']
    for col, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    summary_data = [
        ['Total Employees', 12],
        ['Average Salary', 83667],
        ['Highest Salary', 112000],
        ['Lowest Salary', 62000],
        ['Total Departments', 5],
        ['Average Rating', 4.07],
        ['Employees Above Avg Rating', 7],
        ['Total Payroll', 1004000],
        ['Report Date', '2025-09-15'],
        ['Fiscal Year', 'FY2025'],
    ]

    for r, row_data in enumerate(summary_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_summary.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 2 and isinstance(val, int) and val > 1000:
                cell.number_format = '$#,##0'

    ws_summary.column_dimensions['A'].width = 28
    ws_summary.column_dimensions['B'].width = 16

    # --- Apply workbook structure protection ---
    wb.security.lockStructure = True
    wb.security.workbookPassword = 'oldstruct'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
