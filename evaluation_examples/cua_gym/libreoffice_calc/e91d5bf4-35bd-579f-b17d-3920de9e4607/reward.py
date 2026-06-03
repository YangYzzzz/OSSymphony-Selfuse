"""
Reward Script: Remove document structure protection from workbook
Task ID: calc_ps_016
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): lockStructure is explicitly "0" or False/absent in XML
  Component 2 (0.3): No password hash in workbookProtection element
  Component 3 (0.3): All 3 original sheets still exist with data intact + structure unlocked
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_016'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    sec = wb.security

    # Component 1: lockStructure is not True (0.4 points)
    # In initial_env: lockStructure=True (protected)
    # In golden_env: lockStructure=False (unprotected)
    try:
        if sec is not None:
            lock_val = sec.lockStructure
        else:
            lock_val = None

        if lock_val is True:
            print(f"FAIL: Component 1 — lockStructure is True (still protected)")
        else:
            print(f"PASS: Component 1 — lockStructure is {lock_val} (not protected) (0.4 pts)")
            total_score += 0.4
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: workbookPassword is cleared (0.3 points)
    # In initial_env: workbookPassword='E8BD' (hashed value present)
    # In golden_env: workbookPassword=None (no password)
    try:
        if sec is not None:
            wb_pwd = sec.workbookPassword
        else:
            wb_pwd = None

        if wb_pwd is not None and wb_pwd != '':
            print(f"FAIL: Component 2 — workbookPassword is {repr(wb_pwd)} (still has password)")
        else:
            print(f"PASS: Component 2 — workbookPassword is cleared: {repr(wb_pwd)} (0.3 pts)")
            total_score += 0.3
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: All 3 original sheets still exist with data intact (0.3 points)
    # Anchored to task change: only awards if structure IS unlocked
    # (won't score on initial_env where lockStructure=True)
    try:
        expected_sheets = ['Data', 'Charts', 'Summary']
        sheets_ok = all(s in wb.sheetnames for s in expected_sheets)
        data_ok = False

        if sheets_ok:
            ws_data = wb['Data']
            ws_charts = wb['Charts']
            ws_summary = wb['Summary']
            data_ok = (
                ws_data.max_row >= 10
                and ws_charts.max_row >= 3
                and ws_summary.max_row >= 5
            )

        # Anchor: only award when structure protection is actually removed
        structure_unlocked = (sec is None or sec.lockStructure is not True)

        if structure_unlocked and sheets_ok and data_ok:
            print(f"PASS: Component 3 — All 3 sheets present with data intact after unprotect (0.3 pts)")
            total_score += 0.3
        elif not structure_unlocked:
            print(f"FAIL: Component 3 — Structure still protected, cannot award data integrity points")
        elif not sheets_ok:
            print(f"FAIL: Component 3 — Missing sheets. Found: {wb.sheetnames}, expected: {expected_sheets}")
        else:
            print(f"FAIL: Component 3 — Sheet data appears incomplete")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point — no persist hook needed for protection-check tasks
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
