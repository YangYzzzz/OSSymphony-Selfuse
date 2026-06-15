"""
Reward Script: Use INDIRECT to dynamically reference a range on a different sheet
Task ID: calc_lf_026
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): H2 on Summary contains a formula using SUM
  Component 2 (0.3): The formula uses INDIRECT function referencing G2
  Component 3 (0.3): The formula correctly builds a cross-sheet reference to B2:B5
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_026'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Summary sheet must exist
    if 'Summary' not in wb.sheetnames:
        print("FAIL: 'Summary' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Summary']
    h2_value = ws['H2'].value

    # Precondition: H2 must have some value
    if h2_value is None:
        print(f"FAIL: H2 is empty (None)")
        print("REWARD: 0.0")
        return 0.0

    h2_str = str(h2_value).strip()
    h2_upper = h2_str.upper().replace(" ", "")

    print(f"INFO: H2 raw value: {repr(h2_value)}")
    print(f"INFO: H2 normalized: {h2_upper}")

    # Component 1: H2 contains a SUM formula (0.4 points)
    # This checks that the cell has a formula with SUM - the core aggregation requirement
    try:
        if h2_str.startswith('=') and 'SUM' in h2_upper:
            print(f"PASS: Component 1 - H2 contains a SUM formula (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 - H2 does not contain a SUM formula. Found: {h2_str}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: The formula uses INDIRECT and references G2 (0.3 points)
    # This verifies the dynamic referencing mechanism via INDIRECT linked to G2
    try:
        if 'INDIRECT' in h2_upper and 'G2' in h2_upper:
            print(f"PASS: Component 2 - Formula uses INDIRECT referencing G2 (0.3 pts)")
            total_score += 0.3
        else:
            if 'INDIRECT' not in h2_upper:
                print(f"FAIL: Component 2 - Formula does not use INDIRECT. Found: {h2_str}")
            else:
                print(f"FAIL: Component 2 - Formula uses INDIRECT but doesn't reference G2. Found: {h2_str}")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: The INDIRECT construct references B2:B5 range (0.3 points)
    # This verifies the correct target range on the referenced sheet
    try:
        # Accept various valid forms of referencing B2:B5 via INDIRECT
        # e.g., INDIRECT(G2&".B2:B5"), INDIRECT(G2&"!B2:B5"), INDIRECT("'"&G2&"'.B2:B5"), etc.
        # The key requirement is B2:B5 appears in the INDIRECT expression
        if 'B2:B5' in h2_upper:
            print(f"PASS: Component 3 - Formula references B2:B5 range (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 3 - Formula does not reference B2:B5. Found: {h2_str}")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
