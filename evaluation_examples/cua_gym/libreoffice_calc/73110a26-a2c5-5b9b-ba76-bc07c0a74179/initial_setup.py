"""
Initial Setup: INDIRECT dynamic sheet reference with SUM
Task ID: calc_lf_026
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_026'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Sales ---
    ws_sales = wb.active
    ws_sales.title = 'Sales'

    # Headers
    ws_sales['A1'] = 'Product'
    ws_sales['B1'] = 'Revenue'

    # Data rows
    data = [
        ('A', 1000),
        ('B', 2000),
        ('C', 1500),
        ('D', 2500),
    ]
    for r, (product, revenue) in enumerate(data, 2):
        ws_sales.cell(row=r, column=1, value=product)
        ws_sales.cell(row=r, column=2, value=revenue)

    # Add some additional context rows to make the spreadsheet more realistic
    ws_sales['A7'] = 'Total'
    ws_sales['D1'] = 'Region'
    ws_sales['D2'] = 'North'
    ws_sales['D3'] = 'South'
    ws_sales['D4'] = 'East'
    ws_sales['D5'] = 'West'
    ws_sales['C1'] = 'Units Sold'
    ws_sales['C2'] = 150
    ws_sales['C3'] = 230
    ws_sales['C4'] = 180
    ws_sales['C5'] = 310

    # Column widths
    ws_sales.column_dimensions['A'].width = 12
    ws_sales.column_dimensions['B'].width = 14
    ws_sales.column_dimensions['C'].width = 14
    ws_sales.column_dimensions['D'].width = 12

    # --- Sheet 2: Summary ---
    ws_summary = wb.create_sheet('Summary')

    ws_summary['G1'] = 'Sheet Name'
    ws_summary['G2'] = 'Sales'
    ws_summary['H1'] = 'Total Revenue'
    # H2 is intentionally left EMPTY — the agent must fill this in

    # Add some additional context to Summary
    ws_summary['A1'] = 'Report Summary'
    ws_summary['A3'] = 'This sheet aggregates data from referenced sheets.'
    ws_summary['A4'] = 'Use INDIRECT to dynamically pull totals.'

    ws_summary.column_dimensions['G'].width = 16
    ws_summary.column_dimensions['H'].width = 18
    ws_summary.column_dimensions['A'].width = 45

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
