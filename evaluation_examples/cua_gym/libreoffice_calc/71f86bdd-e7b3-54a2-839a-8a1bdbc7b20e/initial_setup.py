"""
Initial Setup: Sales data for pivot table task
Task ID: calc_pivot_011
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_011'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def generate_revenue_rows(target_sum, count, min_val=150, max_val=1200):
    """Generate `count` revenue values that sum exactly to target_sum."""
    # Generate random values, then adjust last one
    random.seed(42)  # reproducible
    values = [random.randint(min_val, max_val) for _ in range(count - 1)]
    last_val = target_sum - sum(values)
    # Ensure last value is reasonable; if not, redistribute
    if last_val < min_val or last_val > max_val * 2:
        # Scale all values proportionally
        raw_sum = sum(values)
        values = [int(v * (target_sum - min_val) / raw_sum) for v in values]
        last_val = target_sum - sum(values)
    values.append(last_val)
    return values


def create_initial():
    random.seed(12345)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sales'

    # Headers
    headers = ['ID', 'Date', 'Region', 'Product', 'Quantity', 'Revenue']
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12

    # Pivot target values (revenue sums per region/product combo)
    pivot_targets = {
        ('North', 'Widget'): 12500,
        ('North', 'Gadget'): 13400,
        ('North', 'Gizmo'): 12100,
        ('South', 'Widget'): 11300,
        ('South', 'Gadget'): 9800,
        ('South', 'Gizmo'): 13500,
        ('East', 'Widget'): 11800,
        ('East', 'Gadget'): 12700,
        ('East', 'Gizmo'): 11200,
        ('West', 'Widget'): 7600,
        ('West', 'Gadget'): 12500,
        ('West', 'Gizmo'): 14400,
    }

    regions = ['North', 'South', 'East', 'West']
    products = ['Widget', 'Gadget', 'Gizmo']

    # Generate 20 rows per combo = 240 rows
    rows_per_combo = 20
    all_rows = []

    for region in regions:
        for product in products:
            target = pivot_targets[(region, product)]
            # Generate revenues that sum to target
            revenues = []
            remaining = target
            for i in range(rows_per_combo):
                if i == rows_per_combo - 1:
                    rev = remaining
                else:
                    avg = remaining / (rows_per_combo - i)
                    rev = random.randint(max(100, int(avg * 0.5)), int(avg * 1.5))
                    rev = min(rev, remaining - (rows_per_combo - i - 1) * 100)
                    rev = max(rev, 100)
                revenues.append(rev)
                remaining -= rev

            for rev in revenues:
                qty = random.randint(1, 50)
                # Generate a date in 2025
                month = random.randint(1, 12)
                day = random.randint(1, 28)
                date_str = f'2025-{month:02d}-{day:02d}'
                all_rows.append([date_str, region, product, qty, rev])

    # Shuffle to make it look realistic (not grouped by region/product)
    random.shuffle(all_rows)

    # Write data
    for idx, row_data in enumerate(all_rows):
        r = idx + 2  # row 2 onwards
        ws.cell(row=r, column=1, value=idx + 1)  # ID
        ws.cell(row=r, column=2, value=row_data[0])  # Date
        ws.cell(row=r, column=3, value=row_data[1])  # Region
        ws.cell(row=r, column=4, value=row_data[2])  # Product
        ws.cell(row=r, column=5, value=row_data[3])  # Quantity
        ws.cell(row=r, column=6, value=row_data[4])  # Revenue

    # Format revenue column as currency
    for r in range(2, 242):
        ws.cell(row=r, column=6).number_format = '#,##0'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Verify sums
    from collections import defaultdict
    sums = defaultdict(int)
    for row_data in all_rows:
        sums[(row_data[1], row_data[2])] += row_data[4]
    for k, v in sorted(sums.items()):
        print(f'  {k}: {v} (expected {pivot_targets[k]})')
    print(f'  Grand total: {sum(sums.values())} (expected 142800)')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
