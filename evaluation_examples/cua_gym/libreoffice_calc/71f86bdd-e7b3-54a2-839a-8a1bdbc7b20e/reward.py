"""
Reward Script: Create a pivot table with regions as rows and products as columns, showing SUM of Revenue
Task ID: calc_pivot_011
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20): A pivot/summary sheet exists beyond the original 'Sales' sheet
  Component 2 (0.20): Pivot table has correct structure (regions as rows, products as columns)
  Component 3 (0.30): Key intersection values match expected (North/Widget, South/Gadget, East/Gizmo, West/Widget)
  Component 4 (0.15): Grand Total row with correct overall total (142800)
  Component 5 (0.15): Row totals (Grand Total column) are correct for each region
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_011'

# Expected pivot data from task context
EXPECTED_REGIONS = ['North', 'South', 'East', 'West']
EXPECTED_PRODUCTS = ['Widget', 'Gadget', 'Gizmo']

# Expected corner values: {(region, product): revenue}
EXPECTED_CORNERS = {
    ('North', 'Widget'): 12500,
    ('South', 'Gadget'): 9800,
    ('East', 'Gizmo'): 11200,
    ('West', 'Widget'): 7600,
}

# Expected row totals per region
EXPECTED_ROW_TOTALS = {
    'North': 38000,
    'South': 34600,
    'East': 35700,
    'West': 34500,
}

EXPECTED_GRAND_TOTAL = 142800


def find_pivot_sheet(wb):
    """Find the sheet that contains the pivot table (any sheet other than 'Sales')."""
    for sn in wb.sheetnames:
        if sn.lower() != 'sales':
            return wb[sn]
    return None


def parse_pivot_table(ws):
    """
    Parse the pivot table from a worksheet.
    Returns:
      - header_row: dict mapping product name -> column index
      - data_rows: dict mapping region name -> {product: value, ...}
      - grand_total_row: dict mapping product -> value (or 'Grand Total' -> value)
      - grand_total_col_idx: column index of the Grand Total column
    """
    # Scan all rows to find the header row (contains product names)
    header_row_idx = None
    header_map = {}  # product_name -> col_idx
    grand_total_col = None

    for r in range(1, ws.max_row + 1):
        row_vals = []
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            row_vals.append((c, val))

        # Check if this row contains at least 2 of the expected products
        product_hits = 0
        for col_idx, val in row_vals:
            if val and str(val).strip() in EXPECTED_PRODUCTS:
                product_hits += 1
        if product_hits >= 2:
            header_row_idx = r
            for col_idx, val in row_vals:
                if val and str(val).strip() in EXPECTED_PRODUCTS:
                    header_map[str(val).strip()] = col_idx
                if val and 'grand' in str(val).strip().lower() and 'total' in str(val).strip().lower():
                    grand_total_col = col_idx
            break

    if header_row_idx is None:
        return None, None, None, None

    # Parse data rows (regions)
    data_rows = {}
    grand_total_row = {}

    for r in range(header_row_idx + 1, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        if label is None:
            continue
        label_str = str(label).strip()

        if label_str in EXPECTED_REGIONS:
            row_data = {}
            for prod, col_idx in header_map.items():
                row_data[prod] = ws.cell(row=r, column=col_idx).value
            if grand_total_col:
                row_data['_row_total'] = ws.cell(row=r, column=grand_total_col).value
            data_rows[label_str] = row_data

        elif 'grand' in label_str.lower() and 'total' in label_str.lower():
            for prod, col_idx in header_map.items():
                grand_total_row[prod] = ws.cell(row=r, column=col_idx).value
            if grand_total_col:
                grand_total_row['_overall'] = ws.cell(row=r, column=grand_total_col).value

    return header_map, data_rows, grand_total_row, grand_total_col


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: A pivot/summary sheet exists beyond 'Sales' (0.20 points)
    try:
        pivot_ws = find_pivot_sheet(wb)
        if pivot_ws is not None:
            print(f"PASS: Component 1 — Found pivot sheet '{pivot_ws.title}' (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 — No sheet other than 'Sales' found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    if pivot_ws is None:
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Parse the pivot table structure
    header_map, data_rows, grand_total_row, grand_total_col = parse_pivot_table(pivot_ws)

    # Component 2: Pivot table has correct structure (0.20 points)
    # - At least 3 product columns found in headers
    # - At least 4 region rows found as row labels
    try:
        products_found = len(header_map) if header_map else 0
        regions_found = len(data_rows) if data_rows else 0
        struct_score = 0.0

        if products_found >= 3:
            struct_score += 0.10
            print(f"  PASS: Found {products_found}/3 expected product columns: {list(header_map.keys()) if header_map else []}")
        else:
            print(f"  FAIL: Found {products_found}/3 expected product columns")

        if regions_found >= 4:
            struct_score += 0.10
            print(f"  PASS: Found {regions_found}/4 expected region rows: {list(data_rows.keys()) if data_rows else []}")
        else:
            print(f"  FAIL: Found {regions_found}/4 expected region rows")

        if struct_score > 0:
            print(f"PASS: Component 2 — Structure verified ({struct_score} pts)")
            total_score += struct_score
        else:
            print(f"FAIL: Component 2 — Pivot table structure not found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    if not header_map or not data_rows:
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Component 3: Key intersection values match (0.30 points)
    # Check 4 corner values: North/Widget=12500, South/Gadget=9800, East/Gizmo=11200, West/Widget=7600
    try:
        corner_passes = 0
        corner_total = len(EXPECTED_CORNERS)
        for (region, product), expected_val in EXPECTED_CORNERS.items():
            if region in data_rows and product in data_rows[region]:
                actual = data_rows[region][product]
                if actual is not None:
                    try:
                        if abs(float(actual) - expected_val) <= 1.0:
                            corner_passes += 1
                            print(f"  PASS: {region}/{product} = {actual} (expected {expected_val})")
                        else:
                            print(f"  FAIL: {region}/{product} = {actual} (expected {expected_val})")
                    except (ValueError, TypeError):
                        print(f"  FAIL: {region}/{product} = {actual} (not numeric, expected {expected_val})")
                else:
                    print(f"  FAIL: {region}/{product} = None (expected {expected_val})")
            else:
                print(f"  FAIL: {region}/{product} not found in pivot data")

        corner_score = 0.30 * (corner_passes / corner_total)
        if corner_score > 0:
            print(f"PASS: Component 3 — {corner_passes}/{corner_total} corner values correct ({corner_score:.2f} pts)")
            total_score += corner_score
        else:
            print(f"FAIL: Component 3 — No corner values matched")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Grand Total row with correct overall total = 142800 (0.15 points)
    try:
        if grand_total_row:
            overall = grand_total_row.get('_overall')
            if overall is not None:
                try:
                    if abs(float(overall) - EXPECTED_GRAND_TOTAL) <= 1.0:
                        print(f"PASS: Component 4 — Grand Total = {overall} (expected {EXPECTED_GRAND_TOTAL}) (0.15 pts)")
                        total_score += 0.15
                    else:
                        print(f"FAIL: Component 4 — Grand Total = {overall} (expected {EXPECTED_GRAND_TOTAL})")
                except (ValueError, TypeError):
                    print(f"FAIL: Component 4 — Grand Total = {overall} (not numeric)")
            else:
                print(f"FAIL: Component 4 — Grand Total overall value not found")
        else:
            print(f"FAIL: Component 4 — No Grand Total row found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Row totals correct for each region (0.15 points)
    try:
        if grand_total_col:
            row_total_passes = 0
            row_total_count = len(EXPECTED_ROW_TOTALS)
            for region, expected_total in EXPECTED_ROW_TOTALS.items():
                if region in data_rows and '_row_total' in data_rows[region]:
                    actual = data_rows[region]['_row_total']
                    if actual is not None:
                        try:
                            if abs(float(actual) - expected_total) <= 1.0:
                                row_total_passes += 1
                                print(f"  PASS: {region} row total = {actual} (expected {expected_total})")
                            else:
                                print(f"  FAIL: {region} row total = {actual} (expected {expected_total})")
                        except (ValueError, TypeError):
                            print(f"  FAIL: {region} row total = {actual} (not numeric)")
                    else:
                        print(f"  FAIL: {region} row total = None (expected {expected_total})")
                else:
                    print(f"  FAIL: {region} row total not found")

            row_total_score = 0.15 * (row_total_passes / row_total_count)
            if row_total_score > 0:
                print(f"PASS: Component 5 — {row_total_passes}/{row_total_count} row totals correct ({row_total_score:.2f} pts)")
                total_score += row_total_score
            else:
                print(f"FAIL: Component 5 — No row totals matched")
        else:
            print(f"FAIL: Component 5 — No Grand Total column found in header")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook for LibreOffice
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
