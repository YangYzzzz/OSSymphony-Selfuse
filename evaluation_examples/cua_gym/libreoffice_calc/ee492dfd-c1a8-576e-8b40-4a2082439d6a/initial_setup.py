"""
Initial Setup: Product inventory table with some rows having blank or N/A Supplier values.
Task ID: osworld_calc_hide_rows_na_002
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_hide_rows_na_002'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Inventory ---
    ws = wb.active
    ws.title = 'Inventory'

    # Headers
    headers = ['SKU', 'Product Name', 'Supplier', 'Stock Qty', 'Unit Cost']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Data rows — some rows intentionally have 'N/A' or blank Supplier
    # Rows with blank or N/A supplier: rows 3, 6, 9, 12, 15
    data = [
        # SKU, Product Name, Supplier, Stock Qty, Unit Cost
        ['SKU-1001', 'Wireless Keyboard',     'TechSource Ltd',       120,  29.99],
        ['SKU-1002', 'USB-C Hub 7-Port',      '',                      45,  34.50],
        ['SKU-1003', 'Ergonomic Mouse',       'PeriphPlus Corp',      200,  22.75],
        ['SKU-1004', '27" LED Monitor',       'DisplayMax Inc',        30, 189.00],
        ['SKU-1005', 'HDMI Cable 2m',         'N/A',                  310,   8.99],
        ['SKU-1006', 'Laptop Stand Aluminum', 'ComfortDesk Co',        88,  45.00],
        ['SKU-1007', 'Mechanical Keyboard',   'TechSource Ltd',        65,  79.99],
        ['SKU-1008', 'Webcam 1080p',          '',                      52,  55.00],
        ['SKU-1009', 'USB Microphone',        'SoundGear Pro',        110,  49.00],
        ['SKU-1010', 'Cable Management Kit',  'OfficePro Supply',     250,  12.49],
        ['SKU-1011', 'Desk Lamp LED',         'N/A',                   75,  27.00],
        ['SKU-1012', 'Monitor Arm Dual',      'DeskFlex Systems',      40,  95.00],
        ['SKU-1013', 'Portable SSD 1TB',      'DataStore Solutions',   60, 109.99],
        ['SKU-1014', 'Surge Protector 8-Way', '',                      95,  18.75],
        ['SKU-1015', 'Headphone Stand',       'AudioMount Inc',       130,  16.50],
        ['SKU-1016', 'Screen Cleaning Kit',   'N/A',                  400,   6.25],
        ['SKU-1017', 'Bluetooth Numpad',      'TechSource Ltd',        77,  24.99],
        ['SKU-1018', 'USB 3.0 Flash Drive',   'DataStore Solutions',  500,   9.99],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 11
    ws.column_dimensions['E'].width = 11

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open LibreOffice Calc with the initial file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
