"""
Reward Script: Hide rows where Supplier column is blank or contains 'N/A'
Task ID: osworld_calc_hide_rows_na_002
Domain: libreoffice_calc
Scoring:
  Component 1: All rows with 'N/A' in Supplier column are hidden (0.4 pts)
  Component 2: All rows with blank/None Supplier column are hidden (0.4 pts)
  Component 3: No valid-supplier rows are incorrectly hidden (0.2 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_hide_rows_na_002'


def verify_task(file_path):
    """
    Verify that rows where Supplier is blank or 'N/A' are hidden.
    The data rows should still be present but hidden (not deleted).
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Access the Inventory sheet
    try:
        ws = wb['Inventory']
    except KeyError:
        # Fall back to active sheet if 'Inventory' sheet not found
        ws = wb.active
        print(f"WARN: 'Inventory' sheet not found, using active sheet: {ws.title}")

    # --- Precondition: verify the file has data with at least 2 rows ---
    if ws.max_row < 2:
        print("CRITICAL: File has fewer than 2 rows — likely corrupted or empty.")
        print("REWARD: 0.0")
        return 0.0

    # Identify which column is 'Supplier'
    # According to task_config context: columns are SKU, Product Name, Supplier, Stock Qty, Unit Cost
    # Supplier is column 3 (C). But we verify dynamically by checking the header row.
    supplier_col = None
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    for idx, val in enumerate(header_row, start=1):
        if val and str(val).strip().lower() == 'supplier':
            supplier_col = idx
            break

    if supplier_col is None:
        print("CRITICAL: Could not find 'Supplier' header column in row 1.")
        print("REWARD: 0.0")
        return 0.0

    print(f"INFO: Supplier column found at column {supplier_col}")

    # Scan all data rows (skip header row 1) and classify by supplier value
    na_rows = []         # rows with 'N/A'
    blank_rows = []      # rows with blank/None
    valid_rows = []      # rows with a valid (non-blank, non-N/A) supplier

    for row_num in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_num, column=supplier_col)
        val = cell.value
        val_stripped = str(val).strip() if val is not None else ''

        if val is None or val_stripped == '':
            blank_rows.append(row_num)
        elif val_stripped.upper() == 'N/A':
            na_rows.append(row_num)
        else:
            valid_rows.append(row_num)

    print(f"INFO: N/A supplier rows: {na_rows}")
    print(f"INFO: Blank supplier rows: {blank_rows}")
    print(f"INFO: Valid supplier rows: {valid_rows}")

    # --- Component 1: All rows with 'N/A' Supplier are hidden (0.4 points) ---
    try:
        if not na_rows:
            print("SKIP: No N/A supplier rows found — cannot evaluate Component 1.")
            # If there are no N/A rows at all, this component is vacuously satisfied
            # but only award points if blank rows are also satisfied (handled separately)
        else:
            na_all_hidden = all(ws.row_dimensions[r].hidden == True for r in na_rows)
            if na_all_hidden:
                print(f"PASS: Component 1 — All {len(na_rows)} 'N/A' supplier rows are hidden: {na_rows} (0.4 pts)")
                total_score += 0.4
            else:
                not_hidden = [r for r in na_rows if ws.row_dimensions[r].hidden != True]
                print(f"FAIL: Component 1 — Expected rows {na_rows} to be hidden, but rows {not_hidden} are NOT hidden")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # --- Component 2: All rows with blank/None Supplier are hidden (0.4 points) ---
    try:
        if not blank_rows:
            print("SKIP: No blank supplier rows found — cannot evaluate Component 2.")
        else:
            blank_all_hidden = all(ws.row_dimensions[r].hidden == True for r in blank_rows)
            if blank_all_hidden:
                print(f"PASS: Component 2 — All {len(blank_rows)} blank supplier rows are hidden: {blank_rows} (0.4 pts)")
                total_score += 0.4
            else:
                not_hidden = [r for r in blank_rows if ws.row_dimensions[r].hidden != True]
                print(f"FAIL: Component 2 — Expected rows {blank_rows} to be hidden, but rows {not_hidden} are NOT hidden")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # --- Component 3: No valid-supplier rows are incorrectly hidden (0.2 points) ---
    # This is a compound check: rows that should NOT be hidden remain visible.
    # It only awards points if combined with actual hiding work (Components 1 or 2 passed).
    try:
        if not valid_rows:
            print("SKIP: No valid supplier rows found — cannot evaluate Component 3.")
        else:
            incorrectly_hidden = [r for r in valid_rows if ws.row_dimensions[r].hidden == True]
            if not incorrectly_hidden:
                # Only award these 0.2 pts if at least one hiding component passed
                # (to avoid awarding for do-nothing initial state)
                if total_score > 0.0:
                    print(f"PASS: Component 3 — No valid supplier rows are incorrectly hidden: {valid_rows} (0.2 pts)")
                    total_score += 0.2
                else:
                    print(f"FAIL: Component 3 — No valid rows hidden (correct), but no hiding was done — not awarding partial credit")
            else:
                print(f"FAIL: Component 3 — Expected rows {incorrectly_hidden} to remain visible, but they are incorrectly hidden")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in a given env
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
