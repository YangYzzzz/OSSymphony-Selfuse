"""
Initial Setup: Create multi-sheet quarterly workbook for PDF export task
Task ID: calc_gsi_054
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_054'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # Style definitions
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    currency_fmt = '$#,##0.00'
    int_fmt = '#,##0'
    pct_fmt = '0.0%'
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # Common column headers for each quarterly sheet
    headers = ["Product Line", "Region", "Units Sold", "Unit Price", "Revenue", "Cost", "Profit Margin"]

    # Quarterly data
    quarterly_data = {
        "Q1 2025": [
            ["Enterprise Software", "North America", 1240, 899.00, 1114760.00, 634500.00, 0.431],
            ["Enterprise Software", "Europe", 870, 899.00, 782130.00, 445200.00, 0.431],
            ["Enterprise Software", "Asia Pacific", 635, 849.00, 539115.00, 330800.00, 0.386],
            ["Cloud Services", "North America", 3450, 149.00, 514050.00, 185700.00, 0.639],
            ["Cloud Services", "Europe", 2180, 149.00, 324820.00, 117400.00, 0.639],
            ["Cloud Services", "Asia Pacific", 1890, 129.00, 243810.00, 107500.00, 0.559],
            ["Hardware", "North America", 560, 2499.00, 1399440.00, 952000.00, 0.320],
            ["Hardware", "Europe", 340, 2499.00, 849660.00, 578000.00, 0.320],
            ["Hardware", "Asia Pacific", 420, 2299.00, 965580.00, 680400.00, 0.296],
            ["Support Plans", "North America", 2800, 59.00, 165200.00, 56000.00, 0.661],
            ["Support Plans", "Europe", 1950, 59.00, 115050.00, 39000.00, 0.661],
            ["Support Plans", "Asia Pacific", 1400, 49.00, 68600.00, 28000.00, 0.592],
        ],
        "Q2 2025": [
            ["Enterprise Software", "North America", 1380, 899.00, 1240620.00, 703800.00, 0.433],
            ["Enterprise Software", "Europe", 920, 899.00, 827080.00, 468400.00, 0.434],
            ["Enterprise Software", "Asia Pacific", 710, 849.00, 602790.00, 369200.00, 0.387],
            ["Cloud Services", "North America", 3890, 149.00, 579610.00, 209300.00, 0.639],
            ["Cloud Services", "Europe", 2540, 149.00, 378460.00, 136700.00, 0.639],
            ["Cloud Services", "Asia Pacific", 2200, 129.00, 283800.00, 125200.00, 0.559],
            ["Hardware", "North America", 480, 2499.00, 1199520.00, 816000.00, 0.320],
            ["Hardware", "Europe", 390, 2499.00, 974610.00, 663000.00, 0.320],
            ["Hardware", "Asia Pacific", 450, 2299.00, 1034550.00, 729000.00, 0.295],
            ["Support Plans", "North America", 3100, 59.00, 182900.00, 62000.00, 0.661],
            ["Support Plans", "Europe", 2200, 59.00, 129800.00, 44000.00, 0.661],
            ["Support Plans", "Asia Pacific", 1650, 49.00, 80850.00, 33000.00, 0.592],
        ],
        "Q3 2025": [
            ["Enterprise Software", "North America", 1150, 929.00, 1068350.00, 598000.00, 0.440],
            ["Enterprise Software", "Europe", 810, 929.00, 752490.00, 421200.00, 0.440],
            ["Enterprise Software", "Asia Pacific", 590, 879.00, 518610.00, 307200.00, 0.408],
            ["Cloud Services", "North America", 4200, 159.00, 667800.00, 226100.00, 0.661],
            ["Cloud Services", "Europe", 2800, 159.00, 445200.00, 150900.00, 0.661],
            ["Cloud Services", "Asia Pacific", 2450, 139.00, 340550.00, 132100.00, 0.612],
            ["Hardware", "North America", 520, 2499.00, 1299480.00, 884000.00, 0.320],
            ["Hardware", "Europe", 310, 2499.00, 774690.00, 527000.00, 0.320],
            ["Hardware", "Asia Pacific", 380, 2299.00, 873620.00, 615600.00, 0.296],
            ["Support Plans", "North America", 3350, 59.00, 197650.00, 67000.00, 0.661],
            ["Support Plans", "Europe", 2400, 59.00, 141600.00, 48000.00, 0.661],
            ["Support Plans", "Asia Pacific", 1800, 49.00, 88200.00, 36000.00, 0.592],
        ],
        "Q4 2025": [
            ["Enterprise Software", "North America", 1520, 929.00, 1412080.00, 790400.00, 0.440],
            ["Enterprise Software", "Europe", 1050, 929.00, 975450.00, 546000.00, 0.440],
            ["Enterprise Software", "Asia Pacific", 780, 879.00, 685620.00, 405600.00, 0.408],
            ["Cloud Services", "North America", 4600, 159.00, 731400.00, 248100.00, 0.661],
            ["Cloud Services", "Europe", 3100, 159.00, 492900.00, 167100.00, 0.661],
            ["Cloud Services", "Asia Pacific", 2750, 139.00, 382250.00, 148300.00, 0.612],
            ["Hardware", "North America", 680, 2499.00, 1699320.00, 1156000.00, 0.320],
            ["Hardware", "Europe", 420, 2499.00, 1049580.00, 714000.00, 0.320],
            ["Hardware", "Asia Pacific", 510, 2299.00, 1172490.00, 826200.00, 0.295],
            ["Support Plans", "North America", 3700, 59.00, 218300.00, 74000.00, 0.661],
            ["Support Plans", "Europe", 2650, 59.00, 156350.00, 53000.00, 0.661],
            ["Support Plans", "Asia Pacific", 2050, 49.00, 100450.00, 41000.00, 0.592],
        ],
    }

    first_sheet = True
    for sheet_name, data in quarterly_data.items():
        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(sheet_name)

        # Write headers
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Write data
        for r, row_data in enumerate(data, 2):
            for c, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = thin_border
                if c == 1:  # Product Line
                    cell.alignment = Alignment(horizontal="left")
                elif c == 2:  # Region
                    cell.alignment = Alignment(horizontal="left")
                elif c == 3:  # Units Sold
                    cell.number_format = int_fmt
                elif c in (4, 5, 6):  # Price, Revenue, Cost
                    cell.number_format = currency_fmt
                elif c == 7:  # Profit Margin
                    cell.number_format = pct_fmt

        # Column widths
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 16

        # Freeze header row
        ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the workbook in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
