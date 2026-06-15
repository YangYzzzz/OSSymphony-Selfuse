"""
Initial Setup: Apply conditional formatting with color scales to visualize salary ranges
Task ID: calc_hr_037
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_037'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'SalaryMap'

    # Headers
    headers = ['Employee', 'Department', 'Salary']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_font_white = Font(name='Calibri', size=11, bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 20 rows of realistic employee data with salaries from 45000 to 150000
    data = [
        ['Sarah Chen', 'Engineering', 125000],
        ['Marcus Johnson', 'Marketing', 68000],
        ['Priya Patel', 'Engineering', 142000],
        ['David Kim', 'Finance', 95000],
        ['Elena Rodriguez', 'Human Resources', 72000],
        ['James O\'Brien', 'Engineering', 150000],
        ['Aisha Mohammed', 'Sales', 82000],
        ['Robert Taylor', 'Finance', 110000],
        ['Lisa Wang', 'Marketing', 65000],
        ['Michael Brown', 'Operations', 78000],
        ['Jennifer Adams', 'Engineering', 135000],
        ['Carlos Gutierrez', 'Sales', 88000],
        ['Natasha Ivanova', 'Human Resources', 58000],
        ['Thomas Wright', 'Operations', 45000],
        ['Sophia Lee', 'Finance', 102000],
        ['Daniel Foster', 'Marketing', 71000],
        ['Rachel Green', 'Sales', 93000],
        ['Ahmed Hassan', 'Engineering', 118000],
        ['Michelle Park', 'Operations', 52000],
        ['Kevin O\'Malley', 'Finance', 87000],
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        ws.cell(row=r, column=2, value=row_data[1])
        salary_cell = ws.cell(row=r, column=3, value=row_data[2])
        salary_cell.number_format = '$#,##0'

    # Set column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14

    # NO conditional formatting in initial state

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
