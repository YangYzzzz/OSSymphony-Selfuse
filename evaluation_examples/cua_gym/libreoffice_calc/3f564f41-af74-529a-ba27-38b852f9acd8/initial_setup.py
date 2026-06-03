"""
Initial Setup: Cross-departmental headcount planning model
Task ID: calc_wf_057
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_057'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

# Styling constants
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)

# Department data: (name, current_hc, monthly_hires, attrition_rate, avg_salary)
DEPARTMENTS = {
    "Engineering": {"current_hc": 145, "monthly_hires": 8, "attrition_rate": 0.03, "avg_salary": 135000},
    "Sales":       {"current_hc": 92,  "monthly_hires": 5, "attrition_rate": 0.05, "avg_salary": 95000},
    "Marketing":   {"current_hc": 48,  "monthly_hires": 3, "attrition_rate": 0.04, "avg_salary": 88000},
    "Operations":  {"current_hc": 67,  "monthly_hires": 4, "attrition_rate": 0.02, "avg_salary": 78000},
}

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = thin_border

def style_data_cell(ws, row, col, number_format=None):
    cell = ws.cell(row=row, column=col)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if number_format:
        cell.number_format = number_format

def create_dept_sheet(wb, dept_name, info, is_first=False):
    """Create a department sheet with headcount projection data (raw values only)."""
    if is_first:
        ws = wb.active
        ws.title = dept_name
    else:
        ws = wb.create_sheet(dept_name)

    # Row 1: Department info header
    ws.cell(row=1, column=1, value=f"{dept_name} Department - Headcount Plan 2026")
    ws.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True, color="2F5496")
    ws.merge_cells("A1:F1")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    # Row 3: Key parameters
    ws.cell(row=3, column=1, value="Parameter")
    ws.cell(row=3, column=2, value="Value")
    style_header_row(ws, 3, 2)

    params = [
        ("Current Headcount", info["current_hc"]),
        ("Planned Monthly Hires", info["monthly_hires"]),
        ("Monthly Attrition Rate", info["attrition_rate"]),
        ("Average Annual Salary", info["avg_salary"]),
    ]
    for i, (param, val) in enumerate(params, 4):
        ws.cell(row=i, column=1, value=param)
        ws.cell(row=i, column=1).border = thin_border
        ws.cell(row=i, column=1).font = Font(name="Calibri", size=11)
        c = ws.cell(row=i, column=2, value=val)
        c.border = thin_border
        if param == "Monthly Attrition Rate":
            c.number_format = '0.00%'
        elif param == "Average Annual Salary":
            c.number_format = '$#,##0'
        else:
            c.number_format = '0'

    # Row 9: Monthly projection headers
    headers = ["Month", "Starting HC", "Attrition", "New Hires", "Ending HC", "Monthly Comp Cost"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=9, column=col, value=h)
    style_header_row(ws, 9, len(headers))

    # Rows 10-21: Monthly projections (raw data values, NO formulas)
    hc = info["current_hc"]
    for m_idx, month in enumerate(MONTHS):
        row = 10 + m_idx
        ws.cell(row=row, column=1, value=month)
        style_data_cell(ws, row, 1)

        ws.cell(row=row, column=2, value=hc)
        style_data_cell(ws, row, 2, '0')

        attrition = round(hc * info["attrition_rate"])
        ws.cell(row=row, column=3, value=attrition)
        style_data_cell(ws, row, 3, '0')

        ws.cell(row=row, column=4, value=info["monthly_hires"])
        style_data_cell(ws, row, 4, '0')

        ending_hc = hc - attrition + info["monthly_hires"]
        ws.cell(row=row, column=5, value=ending_hc)
        style_data_cell(ws, row, 5, '0')

        # Monthly compensation cost = ending_hc * avg_salary / 12
        comp = round(ending_hc * info["avg_salary"] / 12, 2)
        ws.cell(row=row, column=6, value=comp)
        style_data_cell(ws, row, 6, '$#,##0.00')

        hc = ending_hc  # carry forward

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 18

    ws.freeze_panes = "A10"
    return ws

def create_summary_sheet(wb):
    """Create Company Summary sheet with headers and raw month labels only.
    NO aggregation formulas, NO chart -- those are the task for the agent."""
    ws = wb.create_sheet("Company Summary")

    ws.cell(row=1, column=1, value="Company Headcount Summary - 2026")
    ws.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True, color="2F5496")
    ws.merge_cells("A1:G1")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    # Headers
    headers = ["Month", "Engineering", "Sales", "Marketing", "Operations",
               "Total Headcount", "Total Compensation"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, len(headers))

    # Month labels only (rows 4-15), NO data -- agent must fill these
    for m_idx, month in enumerate(MONTHS):
        row = 4 + m_idx
        ws.cell(row=row, column=1, value=month)
        style_data_cell(ws, row, 1)
        # Leave columns B-G empty for the agent to fill with formulas

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 18

    ws.freeze_panes = "A4"
    return ws

def create_initial():
    wb = openpyxl.Workbook()

    # Create 4 department sheets
    first = True
    for dept_name, info in DEPARTMENTS.items():
        create_dept_sheet(wb, dept_name, info, is_first=first)
        first = False

    # Create empty summary sheet
    create_summary_sheet(wb)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
