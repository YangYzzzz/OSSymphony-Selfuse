"""
Reward Script: Cross-departmental headcount planning model
Task ID: calc_wf_057
Domain: libreoffice_calc
Scoring:
  Component 1 (0.40): Company Summary B4:E15 contain cross-sheet formulas referencing dept sheets
  Component 2 (0.20): Company Summary F column has Total HC formulas (=B+C+D+E)
  Component 3 (0.15): Company Summary G column has compensation aggregation formulas
  Component 4 (0.25): Stacked area chart on Company Summary with 4 department series
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_057'

def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: required sheets exist
    required_sheets = ['Engineering', 'Sales', 'Marketing', 'Operations', 'Company Summary']
    for s in required_sheets:
        if s not in wb.sheetnames:
            print(f"CRITICAL: Required sheet '{s}' not found. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0

    dept_sheets = ['Engineering', 'Sales', 'Marketing', 'Operations']

    # Component 1: Company Summary B4:E15 contain cross-sheet reference formulas (0.40 points)
    # In the initial file, these cells are all empty (None).
    # In the golden file, they contain formulas like =Engineering!E10, =Sales!E10, etc.
    try:
        ws_summary = wb['Company Summary']
        formula_count = 0
        total_cells = 48  # 12 months * 4 departments (B4:E15)

        for row in range(4, 16):  # rows 4-15 (Jan-Dec)
            for col in range(2, 6):  # columns B-E (4 departments)
                val = ws_summary.cell(row=row, column=col).value
                if isinstance(val, str) and val.startswith('='):
                    # Check it references a department sheet
                    val_upper = val.upper()
                    if any(dept.upper() + '!' in val_upper for dept in dept_sheets):
                        formula_count += 1

        if formula_count >= 44:  # Allow minor tolerance (92% of 48)
            print(f"PASS: Component 1 -- {formula_count}/{total_cells} cross-sheet formulas found (0.40 pts)")
            total_score += 0.40
        elif formula_count > 0:
            ratio = formula_count / total_cells
            partial = round(0.40 * ratio, 4)
            print(f"PARTIAL: Component 1 -- {formula_count}/{total_cells} cross-sheet formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 -- No cross-sheet formulas in Company Summary B4:E15")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Total Headcount column (F4:F15) has SUM or addition formulas (0.20 points)
    # Initial file has empty F4:F15. Golden has =B4+C4+D4+E4 or =SUM(B4:E4).
    try:
        ws_summary = wb['Company Summary']
        total_hc_formulas = 0

        for row in range(4, 16):
            val = ws_summary.cell(row=row, column=6).value  # Column F
            if isinstance(val, str) and val.startswith('='):
                val_upper = val.upper().replace(' ', '')
                # Should reference B,C,D,E columns (either SUM or addition)
                if ('SUM' in val_upper) or ('B' in val_upper and 'C' in val_upper and 'D' in val_upper and 'E' in val_upper):
                    total_hc_formulas += 1

        if total_hc_formulas >= 11:  # Allow 1 missing
            print(f"PASS: Component 2 -- {total_hc_formulas}/12 Total HC formulas found (0.20 pts)")
            total_score += 0.20
        elif total_hc_formulas > 0:
            partial = round(0.20 * (total_hc_formulas / 12), 4)
            print(f"PARTIAL: Component 2 -- {total_hc_formulas}/12 Total HC formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- No Total HC formulas in F4:F15")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Total Compensation column (G4:G15) has aggregation formulas (0.15 points)
    # Initial file has empty G4:G15. Golden references dept compensation columns.
    try:
        ws_summary = wb['Company Summary']
        comp_formulas = 0

        for row in range(4, 16):
            val = ws_summary.cell(row=row, column=7).value  # Column G
            if isinstance(val, str) and val.startswith('='):
                # Should reference department sheet compensation data
                val_upper = val.upper()
                if any(dept.upper() + '!' in val_upper for dept in dept_sheets):
                    comp_formulas += 1

        if comp_formulas >= 11:
            print(f"PASS: Component 3 -- {comp_formulas}/12 compensation formulas found (0.15 pts)")
            total_score += 0.15
        elif comp_formulas > 0:
            partial = round(0.15 * (comp_formulas / 12), 4)
            print(f"PARTIAL: Component 3 -- {comp_formulas}/12 compensation formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 -- No compensation formulas in G4:G15")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Stacked area chart on Company Summary with 4 department series (0.25 points)
    # Initial file has 0 charts. Golden has 1 AreaChart(grouping=stacked) with 4 series.
    try:
        ws_summary = wb['Company Summary']
        charts = ws_summary._charts

        if len(charts) == 0:
            print(f"FAIL: Component 4 -- No charts found on Company Summary")
        else:
            # Find an area chart among available charts
            area_charts = [c for c in charts if 'area' in c.__class__.__name__.lower()]

            if len(area_charts) > 0:
                chart = area_charts[0]
                class_name = chart.__class__.__name__
                grouping = getattr(chart, 'grouping', None)
                series_count = len(chart.series)

                # Area chart exists (0.08 pts)
                if 'area' in class_name.lower():
                    total_score += 0.08
                    print(f"  PASS: Area chart found (type={class_name})")

                # Chart is stacked (0.07 pts)
                if grouping == 'stacked':
                    total_score += 0.07
                    print(f"  PASS: Chart grouping is stacked")
                else:
                    print(f"  FAIL: Chart grouping is '{grouping}', expected 'stacked'")

                # Chart has >=4 series for 4 departments (0.10 pts)
                if series_count >= 4:
                    total_score += 0.10
                    print(f"  PASS: Chart has {series_count} series (>=4)")
                elif series_count > 0:
                    partial_series = round(0.10 * (series_count / 4), 4)
                    total_score += partial_series
                    print(f"  PARTIAL: Chart has {series_count} series, expected >=4 ({partial_series} pts)")
                else:
                    print(f"  FAIL: Chart has 0 series")
            else:
                # No area chart found -- check if there's any chart (partial credit)
                chart = charts[0]
                series_count = len(chart.series)
                print(f"PARTIAL: Component 4 -- Chart found but type={chart.__class__.__name__}, not AreaChart")
                if series_count >= 4:
                    total_score += 0.05
                    print(f"  Partial credit: has {series_count} series (0.05 pts)")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
