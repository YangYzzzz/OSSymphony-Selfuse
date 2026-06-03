"""
Initial Setup: Create transaction spreadsheet for CSV export task
Task ID: calc_gsi_043
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_043'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Transactions ---
    ws = wb.active
    ws.title = 'Transactions'

    headers = ['Transaction ID', 'Date', 'Customer', 'Description', 'Amount', 'Category', 'Status']
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Transaction data - realistic financial records
    data = [
        ['TXN-2025-0001', '2025-01-03', 'Meridian Solutions LLC', 'Consulting services - Q1 strategy', 4500.00, 'Professional Services', 'Completed'],
        ['TXN-2025-0002', '2025-01-07', 'Chen & Associates', 'Legal review of vendor contract', 2750.00, 'Legal', 'Completed'],
        ['TXN-2025-0003', '2025-01-12', 'TechForward Inc.', 'Cloud hosting - January', 1890.50, 'IT Infrastructure', 'Completed'],
        ['TXN-2025-0004', '2025-01-15', 'GreenLeaf Supplies', 'Office supplies - bulk order', 623.75, 'Office Supplies', 'Completed'],
        ['TXN-2025-0005', '2025-01-18', 'Apex Marketing Group', 'Digital campaign management', 8200.00, 'Marketing', 'Completed'],
        ['TXN-2025-0006', '2025-01-22', 'Rivera Logistics', 'Freight shipping - warehouse transfer', 3150.00, 'Logistics', 'Pending'],
        ['TXN-2025-0007', '2025-01-25', 'Pinnacle Insurance Corp', 'Quarterly liability premium', 5670.00, 'Insurance', 'Completed'],
        ['TXN-2025-0008', '2025-02-01', 'DataStream Analytics', 'BI dashboard license renewal', 4200.00, 'Software', 'Completed'],
        ['TXN-2025-0009', '2025-02-05', 'Hoffman & Partners', 'Annual audit preparation', 9500.00, 'Accounting', 'In Progress'],
        ['TXN-2025-0010', '2025-02-08', 'Cascade Cleaning Services', 'Monthly janitorial - February', 1450.00, 'Facilities', 'Completed'],
        ['TXN-2025-0011', '2025-02-12', 'TechForward Inc.', 'Cloud hosting - February', 1890.50, 'IT Infrastructure', 'Completed'],
        ['TXN-2025-0012', '2025-02-15', 'BlueSky Catering', 'Team lunch event - product launch', 875.25, 'Events', 'Completed'],
        ['TXN-2025-0013', '2025-02-19', 'Meridian Solutions LLC', 'Follow-up consulting - implementation', 6300.00, 'Professional Services', 'Pending'],
        ['TXN-2025-0014', '2025-02-22', 'NovaTech Security', 'Penetration testing - annual review', 7800.00, 'IT Security', 'In Progress'],
        ['TXN-2025-0015', '2025-02-28', 'GreenLeaf Supplies', 'Printer toner and paper restock', 412.90, 'Office Supplies', 'Completed'],
        ['TXN-2025-0016', '2025-03-03', 'Apex Marketing Group', 'Social media ads - March campaign', 5500.00, 'Marketing', 'Pending'],
        ['TXN-2025-0017', '2025-03-07', 'Chen & Associates', 'Employment contract amendments', 1800.00, 'Legal', 'Completed'],
        ['TXN-2025-0018', '2025-03-10', 'Rivera Logistics', 'Express delivery - client materials', 980.00, 'Logistics', 'Completed'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 5:  # Amount column
                cell.number_format = '$#,##0.00'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 42
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    # --- Sheet 2: Summary ---
    ws2 = wb.create_sheet('Summary')
    ws2['A1'] = 'Category'
    ws2['B1'] = 'Total Transactions'
    ws2['C1'] = 'Total Amount'
    for col in range(1, 4):
        cell = ws2.cell(row=1, column=col)
        cell.font = Font(bold=True)

    categories = [
        ('Professional Services', 2, 10800.00),
        ('Legal', 2, 4550.00),
        ('IT Infrastructure', 2, 3781.00),
        ('Office Supplies', 2, 1036.65),
        ('Marketing', 2, 13700.00),
        ('Logistics', 2, 4130.00),
        ('Insurance', 1, 5670.00),
        ('Software', 1, 4200.00),
        ('Accounting', 1, 9500.00),
        ('Facilities', 1, 1450.00),
        ('Events', 1, 875.25),
        ('IT Security', 1, 7800.00),
    ]
    for r, (cat, cnt, amt) in enumerate(categories, 2):
        ws2.cell(row=r, column=1, value=cat)
        ws2.cell(row=r, column=2, value=cnt)
        cell = ws2.cell(row=r, column=3, value=amt)
        cell.number_format = '$#,##0.00'

    ws2.column_dimensions['A'].width = 24
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup - open the spreadsheet in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
