"""
FINAL REWARD SCRIPT - SUCCESS
Task: Summarize the total downloads for each platform in a new sheet with platform names as column headers using the Pivot Table feature.
Generated: 2025-11-24 07:48:11
Status: success
Model: o3
Total Steps: 3
"""

import openpyxl
import math
import os


def verify_pivot_summary(file_path: str) -> float:
    """Verify that the workbook contains a pivot-style summary sheet
    which lists total downloads for each platform as **column headers**.

    Scoring (progressive):
      0.4 – Pivot/summary sheet exists (different from raw data sheet)
      0.2 – Header row contains *all* platform names found in raw data
      0.2 – Totals row (labelled *Sum* or *Total*) exists
      0.2 – Values in that totals row match the sums computed from raw data
      -----
      1.0 – Full compliance with every requirement
    """
    max_score = 1.0
    score = 0.0

    print(f"Loading workbook {file_path} …")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"✗ Unable to load workbook: {e}")
        return 0.0

    print("Workbook sheets:", wb.sheetnames)

    # 1. Locate raw data sheet (must start with the four expected headers)
    data_sheet = None
    for name in wb.sheetnames:
        sh = wb[name]
        try:
            headers = list(next(sh.iter_rows(values_only=True, max_row=1)))
        except StopIteration:
            continue
        if (len(headers) >= 4 and headers[0] == "Platform" and headers[1] == "App"
                and headers[2] == "Month" and headers[3] == "Downloads"):
            data_sheet = sh
            break

    if not data_sheet:
        print("✗ Raw data sheet with expected headers not found")
        return 0.0

    print(f"✓ Found data sheet: {data_sheet.title}")

    # Compute expected totals per platform from raw data
    totals = {}
    for row in data_sheet.iter_rows(values_only=True, min_row=2):
        platform, _, _, downloads = row[:4]
        if platform is None or downloads is None:
            continue
        try:
            downloads = float(downloads)
        except Exception:
            continue
        totals[platform] = totals.get(platform, 0) + downloads

    print("Computed totals per platform:", totals)

    # 2. Look for a separate sheet whose first row contains platform names in columns
    pivot_sheet = None
    for name in wb.sheetnames:
        if name == data_sheet.title:
            continue  # skip raw data sheet
        sh = wb[name]
        try:
            header_row = list(next(sh.iter_rows(values_only=True, max_row=1)))
        except StopIteration:
            continue
        platform_headers = [v for v in header_row[1:] if isinstance(v, str)]
        if platform_headers and set(totals.keys()).issubset(platform_headers):
            pivot_sheet = sh
            break

    if pivot_sheet is None:
        print("✗ Pivot/summary sheet with platform headers not found")
        return score

    print(f"✓ Found pivot sheet: {pivot_sheet.title}")
    score += 0.4  # pivot sheet present

    # 3. Verify header row contains all platform names
    header_row = list(next(pivot_sheet.iter_rows(values_only=True, max_row=1)))
    platform_headers = [v for v in header_row[1:] if isinstance(v, str)]
    if set(totals.keys()).issubset(platform_headers):
        print("✓ Header row includes all platform names")
        score += 0.2
    else:
        print("✗ Header row is missing some platform names")

    # 4. Find a totals row labelled "Sum" or "Total"
    totals_row = None
    for row in pivot_sheet.iter_rows(values_only=True, min_row=2):
        label = row[0]
        if isinstance(label, str) and any(k in label.lower() for k in ("sum", "total")):
            totals_row = list(row)
            break

    if totals_row is None:
        print("✗ Totals row labelled Sum/Total not found")
        print(f"Current score: {score}")
        return score

    print("✓ Found totals row:", totals_row)
    score += 0.2

    # 5. Verify numeric values match expected sums (tolerance 1e-6)
    values_match = True
    for idx, platform in enumerate(platform_headers, start=1):
        if platform not in totals:
            continue  # platform present in header but not in raw data (allowed)
        expected = totals[platform]
        actual = totals_row[idx]
        if actual is None:
            print(f" ✗ Missing value for {platform}")
            values_match = False
            continue
        try:
            actual_num = float(actual)
        except Exception:
            print(f" ✗ Non-numeric value for {platform}: {actual}")
            values_match = False
            continue
        if abs(actual_num - expected) > 1e-6:
            print(f" ✗ Value mismatch for {platform}: expected {expected}, found {actual_num}")
            values_match = False
        else:
            print(f" ✓ Value correct for {platform}: {actual_num}")

    if values_match:
        score += 0.2
    else:
        print("✗ One or more totals do not match expected values")

    final_score = min(score, max_score)
    print(f"Final score: {final_score}")
    return final_score


if __name__ == "__main__":
    # Path to the workbook produced by the user/agent
    path = "/home/user/summarize_the_total_downloads_for_each_platform_in_a_new_sheet_with_platform_names_as_column_headers.xlsx"
    reward = verify_pivot_summary(path)
    print(f"REWARD: {reward}")
