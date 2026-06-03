"""
Initial Setup: Create spreadsheet with evaluation criteria, scores, and weights.
Task ID: calc_nrv_041
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_041'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Headers
    ws['A1'] = 'Criteria'
    ws['B1'] = 'Score'
    ws['C1'] = 'Weight'

    # 19 rows of evaluation criteria (rows 2-20)
    # Scores range 1-10, weights 0.05-0.30 summing to 1.0
    criteria_data = [
        ('Code Quality',           8, 0.10),
        ('Test Coverage',          7, 0.08),
        ('Documentation',          6, 0.05),
        ('Performance',            9, 0.10),
        ('Security',               7, 0.08),
        ('Scalability',            8, 0.07),
        ('User Experience',        9, 0.06),
        ('Maintainability',        7, 0.05),
        ('Error Handling',         6, 0.05),
        ('API Design',             8, 0.06),
        ('Data Integrity',         9, 0.07),
        ('Deployment Process',     5, 0.04),
        ('Monitoring',             6, 0.04),
        ('Accessibility',          4, 0.03),
        ('Internationalization',   3, 0.03),
        ('Load Testing',           7, 0.05),
        ('Code Review Process',    8, 0.05),
        ('CI/CD Pipeline',         6, 0.04),
        ('Disaster Recovery',      5, 0.05),
    ]
    # Weights sum: 0.10+0.08+0.05+0.10+0.08+0.07+0.06+0.05+0.05+0.06+0.07+0.04+0.04+0.03+0.03+0.05+0.05+0.04+0.05 = 1.00

    for r, (criteria, score, weight) in enumerate(criteria_data, 2):
        ws.cell(row=r, column=1, value=criteria)
        ws.cell(row=r, column=2, value=score)
        ws.cell(row=r, column=3, value=weight)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10

    # NO named ranges in initial state
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
