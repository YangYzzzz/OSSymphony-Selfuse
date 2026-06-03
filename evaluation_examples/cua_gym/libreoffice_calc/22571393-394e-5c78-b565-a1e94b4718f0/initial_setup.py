"""
Initial Setup: Spreadsheet with hardcoded threshold/rate formulas in E2:E201
Task ID: calc_tbl_090
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_090'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    # --- Headers ---
    headers = ["ID", "Rep Name", "Sale Amount", "Region", "Bonus"]
    header_font = Font(bold=True, size=11, name="Calibri")
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, name="Calibri", color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14

    # --- Realistic data ---
    first_names = [
        "Sarah", "Marcus", "Elena", "James", "Priya", "David", "Mei",
        "Carlos", "Aisha", "Robert", "Yuki", "Ahmed", "Lisa", "Wei",
        "Olga", "Michael", "Fatima", "Thomas", "Ananya", "Daniel",
        "Sofia", "Kevin", "Nadia", "Brian", "Hiroshi", "Rachel",
        "Omar", "Jessica", "Leo", "Isabelle", "Ryan", "Mina",
        "Patrick", "Zara", "Victor", "Hannah", "Juan", "Lena",
        "Trevor", "Samira"
    ]
    last_names = [
        "Chen", "Johnson", "Petrova", "Williams", "Sharma", "Kim",
        "Wang", "Rodriguez", "Okafor", "Smith", "Tanaka", "Hassan",
        "Anderson", "Zhang", "Ivanova", "Brown", "Ali", "Mueller",
        "Patel", "Garcia", "Martinez", "Lee", "Khoury", "O'Brien",
        "Sato", "Green", "Farah", "Davis", "Rossi", "Dupont",
        "Taylor", "Park", "Murphy", "Singh", "Cruz", "Moore",
        "Hernandez", "Johansson", "Wright", "Nakamura"
    ]
    regions = ["North", "South", "East", "West", "Central",
               "Northeast", "Southeast", "Northwest", "Southwest", "International"]

    # Generate varied thresholds and rates for the 200 rows
    # Each row gets a unique-ish combination
    thresholds = []
    rates = []
    for i in range(200):
        # Thresholds range from 300 to 1200 in steps of 50
        t = random.choice([300, 350, 400, 450, 500, 550, 600, 650,
                           700, 750, 800, 850, 900, 950, 1000, 1050, 1100, 1150, 1200])
        # Rates: 0.05, 0.08, 0.1, 0.12, 0.15, 0.18, 0.2
        r = random.choice([0.05, 0.08, 0.1, 0.12, 0.15, 0.18, 0.2])
        thresholds.append(t)
        rates.append(r)

    for i in range(200):
        row = i + 2  # rows 2..201

        # Column A: ID
        ws.cell(row=row, column=1, value=1000 + i + 1)

        # Column B: Rep Name
        fn = first_names[i % len(first_names)]
        ln = last_names[(i * 7 + 3) % len(last_names)]
        ws.cell(row=row, column=2, value=f"{fn} {ln}")

        # Column C: Sale Amount (realistic range 100-2500)
        amount = round(random.uniform(100, 2500), 2)
        ws.cell(row=row, column=3, value=amount)
        ws.cell(row=row, column=3).number_format = '#,##0.00'

        # Column D: Region
        ws.cell(row=row, column=4, value=regions[i % len(regions)])

        # Column E: Bonus formula with HARDCODED threshold and rate
        t = thresholds[i]
        r = rates[i]
        formula = f"=IF(C{row}>{t},C{row}*{r},0)"
        ws.cell(row=row, column=5, value=formula)
        ws.cell(row=row, column=5).number_format = '#,##0.00'

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f"Initial file created: {OUTPUT}")

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print("GUI_READY: launched LibreOffice Calc with DISPLAY=:0")


create_initial()
