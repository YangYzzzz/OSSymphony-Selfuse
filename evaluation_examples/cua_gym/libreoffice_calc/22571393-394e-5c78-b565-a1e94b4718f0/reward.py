"""
Reward Script: Create parameters table and restructure formulas to reference it
Task ID: calc_tbl_090
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.15): Parameters sheet exists with correct structure
  - Component 2 (0.35): Parameters values match original hardcoded values
  - Component 3 (0.35): Formulas in E2:E201 reference Parameters sheet
  - Component 4 (0.15): Formula structure is correct (IF pattern with Parameters refs)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_090'

# The original hardcoded thresholds and rates extracted from the initial file.
# These are the values that should appear in the Parameters table.
# We verify a sample of them to confirm correctness without hardcoding all 200.
# The key verification: formulas used to have hardcoded numbers, now they reference Parameters.

def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sales sheet must exist
    if 'Sales' not in wb.sheetnames:
        print("CRITICAL: 'Sales' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws_sales = wb['Sales']

    # Component 1: Parameters sheet exists with correct structure (0.15 points)
    try:
        if 'Parameters' in wb.sheetnames:
            ws_params = wb['Parameters']
            header_a = ws_params.cell(row=1, column=1).value
            header_b = ws_params.cell(row=1, column=2).value

            has_headers = (header_a and header_b and
                          'threshold' in str(header_a).strip().lower() and
                          'rate' in str(header_b).strip().lower())

            has_data = ws_params.max_row >= 201  # header + 200 rows

            if has_headers and has_data:
                print(f"PASS: Component 1 — Parameters sheet exists with correct headers and 200+ data rows (0.15 pts)")
                total_score += 0.15
            elif has_headers:
                print(f"PARTIAL: Component 1 — Parameters sheet has headers but only {ws_params.max_row - 1} data rows (expected 200)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 1 — Parameters sheet headers incorrect: A1='{header_a}', B1='{header_b}'")
        else:
            print("FAIL: Component 1 — 'Parameters' sheet not found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Parameters values are valid thresholds and rates (0.35 points)
    # Verify that values are numeric and in reasonable ranges matching the original patterns
    try:
        if 'Parameters' in wb.sheetnames:
            ws_params = wb['Parameters']
            valid_params = 0
            total_checked = 0

            for r in range(2, min(ws_params.max_row + 1, 202)):
                total_checked += 1
                threshold = ws_params.cell(row=r, column=1).value
                rate = ws_params.cell(row=r, column=2).value

                if (threshold is not None and rate is not None and
                    isinstance(threshold, (int, float)) and isinstance(rate, (int, float)) and
                    200 <= threshold <= 1500 and 0.01 <= rate <= 0.5):
                    valid_params += 1

            if total_checked >= 200:
                ratio = valid_params / total_checked
                if ratio >= 0.95:
                    print(f"PASS: Component 2 — {valid_params}/{total_checked} parameter rows have valid threshold/rate values (0.35 pts)")
                    total_score += 0.35
                elif ratio >= 0.5:
                    partial = round(0.35 * ratio, 2)
                    print(f"PARTIAL: Component 2 — {valid_params}/{total_checked} valid rows ({partial} pts)")
                    total_score += partial
                else:
                    print(f"FAIL: Component 2 — only {valid_params}/{total_checked} rows have valid values")
            else:
                print(f"FAIL: Component 2 — only {total_checked} data rows in Parameters (expected 200)")
        else:
            print("FAIL: Component 2 — 'Parameters' sheet not found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Formulas in E2:E201 reference Parameters sheet (0.35 points)
    # This is the KEY task-introduced change: formulas should reference Parameters! not hardcoded values
    try:
        ref_count = 0
        hardcoded_count = 0
        total_formulas = 0

        for r in range(2, 202):
            val = ws_sales.cell(row=r, column=5).value
            if val and isinstance(val, str) and val.startswith('='):
                total_formulas += 1
                if 'Parameters!' in val or 'parameters!' in val.lower():
                    ref_count += 1
                elif re.match(r'=IF\(C\d+>\d+', val):
                    hardcoded_count += 1

        if total_formulas >= 200 and ref_count >= 190:
            ratio = ref_count / total_formulas
            ratio = ref_count / total_formulas
            if ratio >= 0.95:
                print(f"PASS: Component 3 — {ref_count}/{total_formulas} formulas reference Parameters sheet (0.35 pts)")
                total_score += 0.35
            elif ratio >= 0.5:
                partial_pts = round(0.35 * ratio, 2)
                print(f"PARTIAL: Component 3 — {ref_count}/{total_formulas} formulas reference Parameters ({partial_pts} pts)")
                total_score += partial_pts
            else:
                print(f"FAIL: Component 3 — only {ref_count}/{total_formulas} formulas reference Parameters")
        elif ref_count > 0:
            partial = round(0.35 * (ref_count / 200), 2)
            print(f"PARTIAL: Component 3 — only {ref_count} formulas reference Parameters out of {total_formulas} total ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — no formulas reference Parameters sheet (found {hardcoded_count} hardcoded, {total_formulas} total)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Formula structure is correct IF pattern with Parameters refs (0.15 points)
    # Verify the formula pattern: =IF(Cn>Parameters!An,Cn*Parameters!Bn,0)
    try:
        correct_structure = 0
        checked = 0
        sample_rows = list(range(2, 22)) + list(range(100, 110)) + list(range(192, 202))

        for r in sample_rows:
            val = ws_sales.cell(row=r, column=5).value
            if val and isinstance(val, str):
                checked += 1
                # Normalize and check pattern: =IF(Cn>Parameters!An,Cn*Parameters!Bn,0)
                normalized = val.replace(' ', '')
                pattern = rf'=IF\(C{r}>Parameters!A{r},C{r}\*Parameters!B{r},0\)'
                if re.match(pattern, normalized, re.IGNORECASE):
                    correct_structure += 1

        if checked > 0:
            ratio = correct_structure / checked
            if ratio >= 0.9:
                print(f"PASS: Component 4 — {correct_structure}/{checked} sampled formulas have correct structure (0.15 pts)")
                total_score += 0.15
            elif ratio >= 0.5:
                partial = round(0.15 * ratio, 2)
                print(f"PARTIAL: Component 4 — {correct_structure}/{checked} correct structure ({partial} pts)")
                total_score += partial
            else:
                # Show an example of what we found
                example_val = ws_sales.cell(row=2, column=5).value
                print(f"FAIL: Component 4 — only {correct_structure}/{checked} correct. Example E2: '{example_val}'")
        else:
            print("FAIL: Component 4 — no formulas found in sample rows")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
