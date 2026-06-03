"""
Initial Setup: Add formula validation to ProjectHours spreadsheet
Task ID: calc_dop_validate_formula_067
Domain: libreoffice_calc

Creates a ProjectHours spreadsheet with task records and intentionally
some invalid Actual Hours values. No validation exists on column D yet.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_dop_validate_formula_067'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: ProjectHours ---
    ws = wb.active
    ws.title = 'ProjectHours'

    # Headers row 1
    headers = ['Task', 'Assignee', 'Planned Hours', 'Actual Hours', 'Variance']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, name='Calibri', size=11, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 14
    ws.row_dimensions[1].height = 22

    # Task data - 99 rows (rows 2-100)
    # Some Actual Hours deliberately invalid (< Planned or > 2x Planned)
    tasks_data = [
        ('Website Homepage Redesign', 'Sarah Chen', 20),
        ('Database Schema Migration', 'Marcus Johnson', 35),
        ('API Authentication Module', 'Priya Patel', 25),
        ('Mobile App Login Flow', 'James Rivera', 15),
        ('Quarterly Report Automation', 'Emma Liu', 10),
        ('CI/CD Pipeline Setup', 'Noah Williams', 30),
        ('Security Vulnerability Patch', 'Aisha Okafor', 8),
        ('Customer Dashboard Analytics', 'Carlos Mendez', 40),
        ('Email Notification System', 'Yuki Tanaka', 18),
        ('Payment Gateway Integration', 'Sofia Rossi', 28),
        ('Search Engine Optimization', 'Liam O\'Brien', 12),
        ('Cloud Storage Migration', 'Mia Kowalski', 36),
        ('User Permission Management', 'Ethan Park', 22),
        ('Inventory Tracking Module', 'Amara Diallo', 32),
        ('Performance Benchmarking', 'Hugo Schmidt', 14),
        ('SSO Integration', 'Fatima Al-Hassan', 26),
        ('Data Export Feature', 'Lucas Ferreira', 20),
        ('Chatbot Training Pipeline', 'Zoe Thompson', 38),
        ('Load Balancer Configuration', 'Arjun Sharma', 16),
        ('A/B Testing Framework', 'Isabella Garcia', 24),
        ('Backup Recovery System', 'Dmitri Volkov', 33),
        ('Multi-language Support', 'Aiko Nakamura', 9),
        ('Session Management Overhaul', 'Tyler Brooks', 28),
        ('GraphQL API Design', 'Nadia Laurent', 40),
        ('Microservice Deployment', 'Samuel Osei', 35),
        ('Frontend Component Library', 'Chloe Martin', 18),
        ('Caching Strategy Implementation', 'Rafael Costa', 22),
        ('Audit Trail Logging', 'Ingrid Larsson', 13),
        ('Tax Calculation Engine', 'Kwame Asante', 30),
        ('Password Reset Workflow', 'Valentina Cruz', 7),
        ('Real-time Notifications', 'Oliver Hughes', 25),
        ('Data Deduplication Script', 'Rin Suzuki', 11),
        ('Role-based Access Control', 'Adebayo Adewale', 36),
        ('Customer Onboarding Flow', 'Mei Lin Zhang', 20),
        ('Webhook Event Processing', 'Patrick Dumont', 17),
        ('Analytics Dashboard Build', 'Sasha Petrov', 38),
        ('Form Validation Library', 'Tessa Brouwer', 14),
        ('Docker Image Optimization', 'Carlos Lima', 29),
        ('Reporting Engine Upgrade', 'Hana Kimura', 22),
        ('Mobile Push Notifications', 'Darius Edwards', 31),
        ('Rate Limiting Implementation', 'Elena Sorokina', 8),
        ('CSV Bulk Import Tool', 'Ahmad Karim', 19),
        ('Kubernetes Pod Autoscaling', 'Nina Johansson', 40),
        ('Transactional Email Templates', 'George Mensah', 15),
        ('Internal Wiki Migration', 'Lena Hoffmann', 27),
        ('User Profile Edit Flow', 'Mateo Reyes', 12),
        ('Code Review Automation', 'Blessing Eze', 34),
        ('Order Fulfillment Tracker', 'Aleksei Morozov', 21),
        ('Localization Pipeline', 'Yuna Park', 28),
        ('Feature Flag System', 'Benjamin Dupont', 16),
        ('Legacy API Deprecation', 'Chioma Ike', 39),
        ('Geolocation Services Setup', 'Andres Torres', 24),
        ('Scheduled Jobs Manager', 'Lila Petersen', 18),
        ('Contract Management Portal', 'Marcus Boateng', 33),
        ('Feedback Collection Form', 'Sakura Yamamoto', 10),
        ('Resource Allocation Tool', 'Viktor Popov', 26),
        ('Incident Response Playbook', 'Nia Nkrumah', 22),
        ('API Rate Monitor', 'Jian Wei', 14),
        ('Product Catalog Importer', 'Amelia Santos', 37),
        ('Customer Churn Predictor', 'Finn Andersen', 29),
        ('Subscription Billing Engine', 'Kofi Asiedu', 40),
        ('UI Accessibility Audit', 'Vera Kozlov', 11),
        ('Workflow Automation Builder', 'Daniyar Seitkali', 32),
        ('API Documentation Generator', 'Esther Okonkwo', 19),
        ('Smart Filter Component', 'Louis Beaumont', 23),
        ('Data Retention Policy Tool', 'Ingeborg Haugen', 17),
        ('Sales Pipeline Dashboard', 'Esteban Morales', 35),
        ('Customer Support Ticketing', 'Sayuri Watanabe', 28),
        ('Cross-Platform Build System', 'Emmanuel Adjei', 13),
        ('Dark Mode Implementation', 'Agneta Lindqvist', 9),
        ('Predictive Analytics Module', 'Hamza Chaudhry', 38),
        ('OAuth Provider Integration', 'Olga Smirnova', 21),
        ('Test Coverage Expansion', 'Kwabena Ofori', 30),
        ('Notification Preference Center', 'Paola Russo', 16),
        ('Bulk Email Campaign Tool', 'Mikael Eriksson', 24),
        ('Shipping Rate Calculator', 'Aminata Balde', 20),
        ('Dependency Vulnerability Scan', 'Hideo Tanaka', 36),
        ('Multi-factor Authentication', 'Rosie Chambers', 27),
        ('Invoice Generation System', 'Femi Adeyemi', 33),
        ('Customer Lifetime Value Model', 'Irene Novak', 15),
        ('Project Budget Tracker', 'Theo Bergmann', 25),
        ('HR Onboarding Checklist', 'Siti Rahayu', 8),
        ('Document E-Signature Integration', 'Jordan Hayes', 31),
        ('Log Aggregation Platform', 'Petra Vasiliev', 22),
        ('Smart Recommendation Engine', 'Ali Hassan', 40),
        ('Data Quality Dashboard', 'Claire Fontaine', 18),
        ('Application Performance Monitor', 'Nnamdi Obi', 29),
        ('Rebate Calculation Engine', 'Eva Schulz', 34),
        ('PDF Export Functionality', 'Takeshi Mori', 12),
        ('Session Replay Integration', 'Samira Hadj', 26),
        ('Internal Hackathon Tracker', 'Dennis Otieno', 20),
        ('Supplier Portal Launch', 'Yevgenia Bilyk', 37),
        ('Asset Management System', 'Robert Dlamini', 23),
        ('Expense Approval Workflow', 'Mei Xiang', 11),
        ('Employee Directory App', 'Pierre Moreau', 28),
        ('SLA Monitoring Dashboard', 'Funke Adebisi', 32),
        ('Compliance Audit Tool', 'Tunde Balogun', 19),
        ('Network Latency Monitor', 'Astrid Nilsson', 25),
        ('Product Launch Checklist', 'Seun Adesanya', 14),
    ]

    # Determine actual hours: some intentionally invalid
    # For rows 5, 10, 15, 20, 25 (0-indexed 3,8,13,18,23) - make too low (<planned)
    # For rows 8, 16, 24, 32 (0-indexed 7,15,23,31) - make too high (>2x planned)
    # Note: row_index 0-based for tasks_data list
    invalid_low_indices = {3, 9, 14, 20, 26, 33, 41, 48, 55, 62, 70, 77, 84, 91}
    invalid_high_indices = {7, 16, 23, 31, 39, 46, 53, 60, 68, 75, 82, 89}

    for i, (task, assignee, planned) in enumerate(tasks_data):
        row = i + 2  # data rows start at row 2
        actual = planned + 3  # default: slightly over planned but valid (planned <= actual <= 2*planned)

        if i in invalid_low_indices:
            # Actual less than planned (invalid: too low)
            actual = planned - 2
        elif i in invalid_high_indices:
            # Actual more than 2x planned (invalid: too high)
            actual = planned * 2 + 5
        else:
            # Valid: between planned and 2*planned
            import random
            random.seed(i + 42)
            actual = random.randint(planned, planned * 2)

        variance = actual - planned

        ws.cell(row=row, column=1, value=task)
        ws.cell(row=row, column=2, value=assignee)
        ws.cell(row=row, column=3, value=planned)
        ws.cell(row=row, column=4, value=actual)
        ws.cell(row=row, column=5, value=variance)

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet: ProjectHours')
    print(f'Rows: {len(tasks_data)} data rows (rows 2-100)')
    print(f'NO data validation on column D (as required)')


create_initial()
