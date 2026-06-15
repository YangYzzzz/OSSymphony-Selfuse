"""
Reward Script: Make a pivot table showing student enrollment count per course
Task ID: calc_pivot_010
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Pivot table sheet exists (separate from Enrollment)
  Component 2 (0.15): Header row has Course and Count labels
  Component 3 (0.40): All 5 course counts are correct (0.08 each)
  Component 4 (0.20): Grand total row equals 200
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_010'

# Expected course enrollment counts from ground truth
EXPECTED_COUNTS = {
    'BIO150': 35,
    'CS101': 52,
    'ENG102': 45,
    'MATH201': 38,
    'PHY301': 30,
}
EXPECTED_GRAND_TOTAL = 200


def find_pivot_sheet(wb):
    """Find a sheet that looks like a pivot table (not the Enrollment data sheet)."""
    for sn in wb.sheetnames:
        if sn.lower() != 'enrollment':
            ws = wb[sn]
            # Check if this sheet has course-related data (at least 2 rows)
            if ws.max_row >= 2 and ws.max_column >= 2:
                return ws
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: A pivot table sheet exists (separate from Enrollment) (0.25 points)
    try:
        pivot_ws = find_pivot_sheet(wb)
        if pivot_ws is not None:
            print(f"PASS: Component 1 — Pivot table sheet '{pivot_ws.title}' found (0.25 pts)")
            total_score += 0.25
        else:
            print("FAIL: Component 1 — No pivot table sheet found (only 'Enrollment' exists)")
            print(f"  Sheets: {wb.sheetnames}")
            # No pivot sheet means nothing else can pass
            final_score = 0.0
            print(f"\nScore: {final_score}/1.0")
            print(f"REWARD: {final_score}")
            return final_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 2: Header row has Course and Count-related labels (0.15 points)
    try:
        header_a = str(pivot_ws.cell(row=1, column=1).value or '').strip().lower()
        header_b = str(pivot_ws.cell(row=1, column=2).value or '').strip().lower()
        has_course_header = 'course' in header_a
        has_count_header = 'count' in header_b or 'enroll' in header_b or 'number' in header_b or 'total' in header_b or '#' in header_b
        if has_course_header and has_count_header:
            print(f"PASS: Component 2 — Headers found: '{pivot_ws.cell(row=1,column=1).value}', '{pivot_ws.cell(row=1,column=2).value}' (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 — Expected 'Course' and count-related headers, found: '{pivot_ws.cell(row=1,column=1).value}', '{pivot_ws.cell(row=1,column=2).value}'")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: All 5 course counts are correct (0.08 each = 0.40 total)
    # Build a mapping of course -> count from the pivot sheet data rows
    try:
        pivot_data = {}
        for row_idx in range(2, pivot_ws.max_row + 1):
            course_val = pivot_ws.cell(row=row_idx, column=1).value
            count_val = pivot_ws.cell(row=row_idx, column=2).value
            if course_val is not None:
                course_str = str(course_val).strip()
                # Skip grand total rows for this component
                if course_str.lower() in ('grand total', 'total', 'sum'):
                    continue
                try:
                    pivot_data[course_str] = float(count_val) if count_val is not None else None
                except (ValueError, TypeError):
                    pivot_data[course_str] = None

        comp3_score = 0.0
        for course, expected_count in EXPECTED_COUNTS.items():
            if course in pivot_data and pivot_data[course] is not None:
                if abs(pivot_data[course] - expected_count) < 0.5:
                    print(f"PASS: Component 3 — {course} count = {int(pivot_data[course])} (expected {expected_count}) (0.08 pts)")
                    comp3_score += 0.08
                else:
                    print(f"FAIL: Component 3 — {course} count = {pivot_data[course]}, expected {expected_count}")
            else:
                print(f"FAIL: Component 3 — {course} not found in pivot data. Found courses: {list(pivot_data.keys())}")

        total_score += comp3_score
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Grand total row equals 200 (0.20 points)
    try:
        grand_total_found = False
        for row_idx in range(2, pivot_ws.max_row + 1):
            label = pivot_ws.cell(row=row_idx, column=1).value
            value = pivot_ws.cell(row=row_idx, column=2).value
            if label is not None and str(label).strip().lower() in ('grand total', 'total', 'sum'):
                if value is not None:
                    try:
                        total_val = float(value)
                        if abs(total_val - EXPECTED_GRAND_TOTAL) < 0.5:
                            print(f"PASS: Component 4 — Grand total = {int(total_val)} (expected {EXPECTED_GRAND_TOTAL}) (0.20 pts)")
                            total_score += 0.20
                            grand_total_found = True
                        else:
                            print(f"FAIL: Component 4 — Grand total = {total_val}, expected {EXPECTED_GRAND_TOTAL}")
                            grand_total_found = True
                    except (ValueError, TypeError):
                        print(f"FAIL: Component 4 — Grand total value not numeric: {value}")
                        grand_total_found = True
                break
        if not grand_total_found:
            # Also check if sum of all course counts equals 200 even without explicit grand total row
            all_counts = []
            for row_idx in range(2, pivot_ws.max_row + 1):
                v = pivot_ws.cell(row=row_idx, column=2).value
                if v is not None:
                    try:
                        all_counts.append(float(v))
                    except (ValueError, TypeError):
                        pass
            if all_counts and abs(sum(all_counts) - EXPECTED_GRAND_TOTAL) < 0.5:
                print(f"FAIL: Component 4 — No explicit grand total row found (sum of values = {int(sum(all_counts))})")
            else:
                print(f"FAIL: Component 4 — No grand total row found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook for LibreOffice
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
