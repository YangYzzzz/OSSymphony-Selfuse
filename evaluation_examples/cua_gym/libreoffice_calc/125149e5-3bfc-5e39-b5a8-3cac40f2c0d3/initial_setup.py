"""
Initial Setup: Create enrollment spreadsheet for pivot table task
Task ID: calc_pivot_010
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_010'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


# --- Realistic name pools ---
FIRST_NAMES = [
    "Sarah", "Marcus", "Emily", "James", "Priya", "Carlos", "Wei", "Fatima",
    "David", "Olivia", "Raj", "Sofia", "Liam", "Aisha", "Noah", "Yuki",
    "Ethan", "Maria", "Lucas", "Hannah", "Alex", "Zara", "Daniel", "Mei",
    "Ryan", "Chloe", "Ahmed", "Isabella", "Kevin", "Elena", "Tyler", "Nadia",
    "Brandon", "Lily", "Jason", "Amara", "Nathan", "Grace", "Justin", "Sara",
    "Derek", "Anna", "Kyle", "Mia", "Patrick", "Julia", "Sean", "Rachel",
    "Jordan", "Tanya"
]

LAST_NAMES = [
    "Chen", "Johnson", "Patel", "Garcia", "Kim", "Brown", "Nguyen", "Martinez",
    "Anderson", "Lee", "Thompson", "Walker", "Harris", "Robinson", "Clark",
    "Lewis", "Young", "Hall", "Allen", "Wright", "Torres", "Hill", "Scott",
    "Green", "Adams", "Baker", "Rivera", "Mitchell", "Carter", "Phillips",
    "Evans", "Turner", "Parker", "Edwards", "Collins", "Stewart", "Morris",
    "Reed", "Cook", "Morgan", "Bell", "Murphy", "Bailey", "Cooper", "Price",
    "Brooks", "Ross", "Henderson", "Coleman", "Jenkins"
]

SEMESTERS = ["Fall 2024", "Spring 2025", "Fall 2025"]

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enrollment"

    # --- Headers ---
    headers = ["EnrollID", "StudentName", "Course", "Section", "Semester"]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = header_border

    # --- Build course assignments with exact counts ---
    # CS101=52, MATH201=38, ENG102=45, PHY301=30, BIO150=35 => total=200
    course_list = (
        ["CS101"] * 52 +
        ["MATH201"] * 38 +
        ["ENG102"] * 45 +
        ["PHY301"] * 30 +
        ["BIO150"] * 35
    )
    random.shuffle(course_list)

    sections = ["A", "B", "C"]

    # --- Write 200 data rows ---
    for i in range(200):
        row = i + 2
        enroll_id = i + 1
        first = random.choice(FIRST_NAMES)
        last = random.choice(LAST_NAMES)
        student_name = f"{first} {last}"
        course = course_list[i]
        section = random.choice(sections)
        semester = random.choice(SEMESTERS)

        ws.cell(row=row, column=1, value=enroll_id)
        ws.cell(row=row, column=2, value=student_name)
        ws.cell(row=row, column=3, value=course)
        ws.cell(row=row, column=4, value=section)
        ws.cell(row=row, column=5, value=semester)

    # --- Column widths ---
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
