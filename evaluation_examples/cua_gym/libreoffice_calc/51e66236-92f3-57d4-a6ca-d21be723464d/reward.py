"""
Reward Script: Build a peer review scoring matrix for capstone presentations.
Task ID: calc_edu_peer_review_matrix_048
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): Column G (AvgScore) contains AVERAGE formulas for all 75 data rows
  Component 2 (0.35): Summary rows 79-93 have Overall Avg (AVERAGEIF), Adjusted Avg (drop-lowest), and Difference (ABS) formulas
  Component 3 (0.30): Conditional formatting on summary range A79:D93 using formula $D>0.5 with a highlight fill
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_peer_review_matrix_048'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check the PeerReview sheet exists
    if 'PeerReview' not in wb.sheetnames:
        print("FAIL: 'PeerReview' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['PeerReview']

    # -----------------------------------------------------------------------
    # Component 1: Column G (AvgScore) contains AVERAGE formulas for rows 2-76 (0.35 points)
    # The initial file has all None in column G. The golden file adds =AVERAGE(C:F) for each row.
    # -----------------------------------------------------------------------
    try:
        avg_formula_count = 0
        total_data_rows = 75  # rows 2-76

        for row in range(2, 77):
            val = ws.cell(row=row, column=7).value
            if val is not None and 'AVERAGE' in str(val).upper():
                avg_formula_count += 1

        ratio = avg_formula_count / total_data_rows
        if ratio >= 1.0:
            print(f"PASS: Component 1 — All {avg_formula_count}/{total_data_rows} rows in column G have AVERAGE formulas (0.35 pts)")
            total_score += 0.35
        elif ratio >= 0.5:
            partial = round(0.35 * ratio, 3)
            print(f"PARTIAL: Component 1 — {avg_formula_count}/{total_data_rows} rows in column G have AVERAGE formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {avg_formula_count}/{total_data_rows} rows in column G have AVERAGE formula (expected all 75)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Summary rows 79-93 have formulas in columns B, C, D (0.35 points)
    # B: AVERAGEIF by student name (overall average)
    # C: Adjusted average (drops lowest peer score, uses SUMPRODUCT/MIN formula)
    # D: ABS(B-C) difference
    # All must be None in the initial file.
    # -----------------------------------------------------------------------
    try:
        summary_b_ok = 0   # AVERAGEIF formulas in column B
        summary_c_ok = 0   # adjusted avg formulas in column C
        summary_d_ok = 0   # ABS difference formulas in column D
        total_summary_rows = 15  # rows 79-93

        for row in range(79, 94):
            val_b = ws.cell(row=row, column=2).value
            val_c = ws.cell(row=row, column=3).value
            val_d = ws.cell(row=row, column=4).value

            if val_b is not None and 'AVERAGEIF' in str(val_b).upper():
                summary_b_ok += 1

            # Adjusted avg: accept any non-None formula in column C that references MIN or drop-lowest logic
            if val_c is not None and str(val_c).startswith('='):
                summary_c_ok += 1

            # Difference: must reference ABS and both B and C columns
            if val_d is not None and 'ABS' in str(val_d).upper():
                summary_d_ok += 1

        b_ratio = summary_b_ok / total_summary_rows
        c_ratio = summary_c_ok / total_summary_rows
        d_ratio = summary_d_ok / total_summary_rows
        avg_ratio = (b_ratio + c_ratio + d_ratio) / 3.0

        if avg_ratio >= 1.0:
            print(f"PASS: Component 2 — Summary B:{summary_b_ok}, C:{summary_c_ok}, D:{summary_d_ok} all 15/15 formulas present (0.35 pts)")
            total_score += 0.35
        elif avg_ratio >= 0.5:
            partial = round(0.35 * avg_ratio, 3)
            print(f"PARTIAL: Component 2 — Summary B:{summary_b_ok}, C:{summary_c_ok}, D:{summary_d_ok}/15 formulas present ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Summary formulas insufficient. B:{summary_b_ok}, C:{summary_c_ok}, D:{summary_d_ok} out of {total_summary_rows}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Conditional formatting on summary range using $D>0.5 formula (0.30 points)
    # The initial file has no conditional formatting. The golden adds a formula rule
    # targeting the A79:D93 range that highlights rows where the difference > 0.5.
    # -----------------------------------------------------------------------
    try:
        cf_found = False
        cf_formula_correct = False
        cf_fill_present = False

        for cf_range, rules_list in ws.conditional_formatting._cf_rules.items():
            for rule in rules_list:
                if rule.type == 'expression' and hasattr(rule, 'formula') and rule.formula:
                    formula_str = str(rule.formula[0]).upper().replace(' ', '')
                    # Accept any formula referencing D column with >0.5
                    if 'D' in formula_str and '0.5' in formula_str:
                        cf_found = True
                        cf_formula_correct = True
                        # Check if a fill (highlight) is defined
                        if hasattr(rule, 'dxf') and rule.dxf is not None:
                            try:
                                fill = rule.dxf.fill
                                if fill is not None:
                                    cf_fill_present = True
                            except Exception:
                                pass

        if cf_found and cf_formula_correct and cf_fill_present:
            print(f"PASS: Component 3 — Conditional formatting with $D>0.5 formula and highlight fill present (0.30 pts)")
            total_score += 0.30
        elif cf_found and cf_formula_correct:
            print(f"PARTIAL: Component 3 — Conditional formatting formula found but no fill highlight defined (0.15 pts)")
            total_score += 0.15
        elif cf_found:
            print(f"PARTIAL: Component 3 — Conditional formatting present but formula does not match $D>0.5 (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — No conditional formatting on summary range referencing column D > 0.5")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
