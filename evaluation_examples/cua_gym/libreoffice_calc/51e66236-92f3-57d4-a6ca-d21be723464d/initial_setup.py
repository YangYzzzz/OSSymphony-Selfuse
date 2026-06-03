"""
Initial Setup: Peer Review Scoring Matrix for Capstone Presentations
Task ID: calc_edu_peer_review_matrix_048
Domain: libreoffice_calc

Creates a spreadsheet with:
- Sheet 'PeerReview' with 15 students x 5 criteria flat layout
- Peer scores (1-5) from 4 raters, AvgScore column empty (no formulas yet)
- Student summary section (rows 78-93) with empty formula columns
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_peer_review_matrix_048'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    # ── PeerReview sheet ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'PeerReview'

    # ---------- Column widths ----------
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12

    # ---------- Row 1: Headers ----------
    headers = ['Student', 'Criterion', 'Peer1', 'Peer2', 'Peer3', 'Peer4', 'AvgScore']
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF', name='Calibri', size=11)
    thin = Side(style='thin', color='000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = header_border

    ws.row_dimensions[1].height = 20

    # ---------- Student and criterion data ----------
    students = [
        'Aiden Clarke',
        'Brianna Nguyen',
        'Carlos Ortega',
        'Diana Patel',
        'Ethan Morrison',
        'Fatima Hassan',
        'George Yamamoto',
        'Hannah Kowalski',
        'Ivan Petrov',
        'Jasmine Williams',
        'Kevin O\'Brien',
        'Layla Al-Rashid',
        'Marcus Thompson',
        'Nina Johansson',
        'Oscar Fernandez',
    ]

    criteria = [
        'Research Depth',
        'Presentation Clarity',
        'Technical Accuracy',
        'Teamwork & Collaboration',
        'Innovation & Creativity',
    ]

    # Realistic peer score data (1-5), varied per student/criterion/rater
    # Each student has 5 rows (one per criterion)
    # Scores are designed so some students will have adjusted avg differ by >0.5
    raw_scores = {
        'Aiden Clarke': {
            'Research Depth':            [4, 4, 5, 4],
            'Presentation Clarity':      [3, 4, 4, 4],
            'Technical Accuracy':        [5, 4, 4, 5],
            'Teamwork & Collaboration':  [4, 4, 4, 4],
            'Innovation & Creativity':   [3, 4, 5, 4],
        },
        'Brianna Nguyen': {
            'Research Depth':            [5, 5, 5, 2],  # drop-lowest effect: 2 dropped
            'Presentation Clarity':      [5, 5, 4, 2],
            'Technical Accuracy':        [4, 5, 5, 2],
            'Teamwork & Collaboration':  [5, 4, 5, 2],
            'Innovation & Creativity':   [5, 5, 5, 2],
        },
        'Carlos Ortega': {
            'Research Depth':            [3, 3, 4, 3],
            'Presentation Clarity':      [4, 3, 3, 3],
            'Technical Accuracy':        [3, 4, 3, 3],
            'Teamwork & Collaboration':  [3, 3, 4, 3],
            'Innovation & Creativity':   [4, 3, 3, 3],
        },
        'Diana Patel': {
            'Research Depth':            [5, 5, 4, 1],  # drop-lowest effect: 1 dropped
            'Presentation Clarity':      [5, 4, 5, 1],
            'Technical Accuracy':        [4, 5, 5, 1],
            'Teamwork & Collaboration':  [5, 5, 4, 1],
            'Innovation & Creativity':   [4, 5, 5, 1],
        },
        'Ethan Morrison': {
            'Research Depth':            [4, 4, 4, 3],
            'Presentation Clarity':      [3, 4, 4, 4],
            'Technical Accuracy':        [4, 3, 4, 4],
            'Teamwork & Collaboration':  [4, 4, 3, 4],
            'Innovation & Creativity':   [4, 4, 4, 3],
        },
        'Fatima Hassan': {
            'Research Depth':            [5, 4, 5, 5],
            'Presentation Clarity':      [5, 5, 4, 5],
            'Technical Accuracy':        [5, 5, 5, 4],
            'Teamwork & Collaboration':  [4, 5, 5, 5],
            'Innovation & Creativity':   [5, 4, 5, 5],
        },
        'George Yamamoto': {
            'Research Depth':            [2, 3, 3, 2],
            'Presentation Clarity':      [3, 2, 3, 2],
            'Technical Accuracy':        [2, 3, 2, 3],
            'Teamwork & Collaboration':  [3, 2, 3, 2],
            'Innovation & Creativity':   [2, 3, 2, 3],
        },
        'Hannah Kowalski': {
            'Research Depth':            [4, 5, 4, 4],
            'Presentation Clarity':      [4, 4, 5, 4],
            'Technical Accuracy':        [5, 4, 4, 5],
            'Teamwork & Collaboration':  [4, 5, 4, 4],
            'Innovation & Creativity':   [4, 4, 5, 4],
        },
        'Ivan Petrov': {
            'Research Depth':            [3, 3, 3, 1],  # drop-lowest effect: 1 dropped
            'Presentation Clarity':      [3, 3, 3, 1],
            'Technical Accuracy':        [3, 3, 3, 1],
            'Teamwork & Collaboration':  [3, 3, 3, 1],
            'Innovation & Creativity':   [3, 3, 3, 1],
        },
        'Jasmine Williams': {
            'Research Depth':            [4, 5, 4, 4],
            'Presentation Clarity':      [5, 4, 4, 4],
            'Technical Accuracy':        [4, 4, 5, 4],
            'Teamwork & Collaboration':  [4, 4, 4, 5],
            'Innovation & Creativity':   [5, 4, 4, 4],
        },
        "Kevin O'Brien": {
            'Research Depth':            [3, 4, 3, 3],
            'Presentation Clarity':      [3, 3, 4, 3],
            'Technical Accuracy':        [4, 3, 3, 3],
            'Teamwork & Collaboration':  [3, 3, 3, 4],
            'Innovation & Creativity':   [3, 4, 3, 3],
        },
        'Layla Al-Rashid': {
            'Research Depth':            [5, 4, 5, 4],
            'Presentation Clarity':      [4, 5, 5, 4],
            'Technical Accuracy':        [5, 5, 4, 4],
            'Teamwork & Collaboration':  [4, 4, 5, 5],
            'Innovation & Creativity':   [5, 5, 4, 4],
        },
        'Marcus Thompson': {
            'Research Depth':            [4, 4, 4, 4],
            'Presentation Clarity':      [4, 4, 4, 4],
            'Technical Accuracy':        [4, 4, 4, 4],
            'Teamwork & Collaboration':  [4, 4, 4, 4],
            'Innovation & Creativity':   [4, 4, 4, 4],
        },
        'Nina Johansson': {
            'Research Depth':            [5, 5, 5, 1],  # drop-lowest effect: 1 dropped
            'Presentation Clarity':      [5, 5, 5, 1],
            'Technical Accuracy':        [5, 5, 5, 1],
            'Teamwork & Collaboration':  [5, 5, 5, 1],
            'Innovation & Creativity':   [5, 5, 5, 1],
        },
        'Oscar Fernandez': {
            'Research Depth':            [3, 4, 3, 3],
            'Presentation Clarity':      [4, 3, 4, 3],
            'Technical Accuracy':        [3, 3, 4, 3],
            'Teamwork & Collaboration':  [3, 4, 3, 4],
            'Innovation & Creativity':   [4, 3, 3, 3],
        },
    }

    data_fill = PatternFill(fill_type=None)  # no fill
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    row = 2
    for student in students:
        for criterion in criteria:
            scores = raw_scores[student][criterion]
            ws.cell(row=row, column=1, value=student).alignment = left_align
            ws.cell(row=row, column=1).border = cell_border

            ws.cell(row=row, column=2, value=criterion).alignment = left_align
            ws.cell(row=row, column=2).border = cell_border

            for peer_idx, score in enumerate(scores, 3):  # cols C=3, D=4, E=5, F=6
                c = ws.cell(row=row, column=peer_idx, value=score)
                c.alignment = center_align
                c.border = cell_border

            # Column G (AvgScore) — intentionally empty (no formula)
            g_cell = ws.cell(row=row, column=7, value=None)
            g_cell.border = cell_border

            row += 1

    # Row 77: blank separator
    row = 77
    for col in range(1, 8):
        ws.cell(row=row, column=col, value=None)

    # ---------- Student Summary section (rows 78-93) ----------
    # Row 78: summary headers
    summary_headers = ['Student', 'Overall Avg', 'Adjusted Avg', 'Difference']
    summary_fill = PatternFill(start_color='FF70AD47', end_color='FF70AD47', fill_type='solid')
    summary_font = Font(bold=True, color='FFFFFFFF', name='Calibri', size=11)

    for col_idx, h in enumerate(summary_headers, 1):
        cell = ws.cell(row=78, column=col_idx, value=h)
        cell.font = summary_font
        cell.fill = summary_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = cell_border

    ws.row_dimensions[78].height = 20

    # Rows 79-93: one row per student, summary columns empty (no formulas yet)
    for s_idx, student in enumerate(students):
        r = 79 + s_idx
        # Column A: student name
        name_cell = ws.cell(row=r, column=1, value=student)
        name_cell.alignment = left_align
        name_cell.border = cell_border

        # Columns B, C, D: empty (no formulas yet)
        for col in range(2, 5):
            c = ws.cell(row=r, column=col, value=None)
            c.border = cell_border
            c.alignment = center_align

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheets: {wb.sheetnames}')
    print(f'  Data rows (criteria): rows 2-76 (75 rows)')
    print(f'  Summary section: rows 78-93')
    print(f'  NOTE: Column G (AvgScore) is empty; summary formulas are absent')

create_initial()
