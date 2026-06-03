"""
Initial Setup: Apply custom date format 'DDDD, MMMM DD, YYYY' to date cells
Task ID: calc_lf_085
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_085'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Schedule ---
    ws = wb.active
    ws.title = 'Schedule'

    # Headers
    headers = ['Event', 'Date', 'Location', 'Duration (hrs)', 'Instructor']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    white_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows with realistic content
    data = [
        ['Exam 1', datetime(2025, 3, 15), 'Room 204, Science Hall', 3.0, 'Dr. Patricia Nguyen'],
        ['Exam 2', datetime(2025, 4, 7), 'Lecture Hall B', 2.5, 'Prof. James Mitchell'],
        ['Exam 3', datetime(2025, 5, 20), 'Room 118, Arts Building', 2.0, 'Dr. Linda Okafor'],
        ['Midterm Review', datetime(2025, 3, 10), 'Library Study Room 3', 1.5, 'TA Emily Park'],
        ['Final Review', datetime(2025, 5, 15), 'Online (Zoom)', 2.0, 'Prof. James Mitchell'],
        ['Lab Practical', datetime(2025, 4, 22), 'Chemistry Lab 5', 3.5, 'Dr. Patricia Nguyen'],
        ['Presentation Day', datetime(2025, 5, 1), 'Auditorium', 4.0, 'Dr. Linda Okafor'],
        ['Quiz 1', datetime(2025, 2, 14), 'Room 204, Science Hall', 1.0, 'TA Emily Park'],
        ['Quiz 2', datetime(2025, 3, 28), 'Lecture Hall B', 1.0, 'Prof. James Mitchell'],
        ['Group Project Due', datetime(2025, 4, 15), 'Online Submission', 0.0, 'Dr. Patricia Nguyen'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 2:
                # Default date format — NO custom format applied
                cell.number_format = 'yyyy-mm-dd'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 24

    # --- Sheet 2: Course Info ---
    ws2 = wb.create_sheet('Course Info')
    ws2['A1'] = 'Course'
    ws2['B1'] = 'Credits'
    ws2['C1'] = 'Semester'
    ws2['A1'].font = Font(bold=True)
    ws2['B1'].font = Font(bold=True)
    ws2['C1'].font = Font(bold=True)

    course_data = [
        ['CHEM 301 - Organic Chemistry', 4, 'Spring 2025'],
        ['PHYS 202 - Quantum Mechanics', 3, 'Spring 2025'],
        ['MATH 401 - Real Analysis', 3, 'Spring 2025'],
        ['BIO 250 - Genetics', 4, 'Spring 2025'],
    ]
    for r, row_data in enumerate(course_data, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 15

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
