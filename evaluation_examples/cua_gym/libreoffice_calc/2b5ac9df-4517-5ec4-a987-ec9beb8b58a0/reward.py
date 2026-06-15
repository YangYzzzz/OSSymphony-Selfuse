"""
Reward Script: Apply custom date format 'DDDD, MMMM DD, YYYY' to B2:B4
Task ID: calc_lf_085
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): B2 number_format is 'DDDD, MMMM DD, YYYY'
  Component 2 (0.3): B3 number_format is 'DDDD, MMMM DD, YYYY'
  Component 3 (0.3): B4 number_format is 'DDDD, MMMM DD, YYYY'
"""

import os
import datetime

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_085'

# The exact format string the task requires
EXPECTED_FORMAT = 'DDDD, MMMM DD, YYYY'


def normalize_format(fmt):
    """Normalize a number format string for comparison (case-insensitive, strip whitespace)."""
    if not isinstance(fmt, str):
        return ''
    return fmt.strip().upper()


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Schedule' sheet must exist
    if 'Schedule' not in wb.sheetnames:
        print(f"FAIL: 'Schedule' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Schedule']

    # Precondition: B2:B4 must contain datetime values (dates should still be there)
    for coord in ['B2', 'B3', 'B4']:
        val = ws[coord].value
        if not isinstance(val, datetime.datetime):
            print(f"FAIL: {coord} is not a datetime, found {type(val).__name__}: {repr(val)}")
            print("REWARD: 0.0")
            return 0.0

    expected_norm = normalize_format(EXPECTED_FORMAT)

    # Component 1: B2 has the custom date format (0.4 points)
    try:
        fmt_b2 = ws['B2'].number_format
        if normalize_format(fmt_b2) == expected_norm:
            print(f"PASS: Component 1 -- B2 number_format is '{fmt_b2}' (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 -- B2 number_format is '{fmt_b2}', expected '{EXPECTED_FORMAT}'")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: B3 has the custom date format (0.3 points)
    try:
        fmt_b3 = ws['B3'].number_format
        if normalize_format(fmt_b3) == expected_norm:
            print(f"PASS: Component 2 -- B3 number_format is '{fmt_b3}' (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 -- B3 number_format is '{fmt_b3}', expected '{EXPECTED_FORMAT}'")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: B4 has the custom date format (0.3 points)
    try:
        fmt_b4 = ws['B4'].number_format
        if normalize_format(fmt_b4) == expected_norm:
            print(f"PASS: Component 3 -- B4 number_format is '{fmt_b4}' (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 3 -- B4 number_format is '{fmt_b4}', expected '{EXPECTED_FORMAT}'")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = round(min(total_score, 1.0), 1)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
