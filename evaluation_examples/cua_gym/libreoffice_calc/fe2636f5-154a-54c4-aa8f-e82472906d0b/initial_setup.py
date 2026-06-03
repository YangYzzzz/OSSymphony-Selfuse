"""
Initial Setup: Edit hyperlink in cell C3 to change URL
Task ID: calc_cop_hyperlink_005
Domain: libreoffice_calc

Creates a spreadsheet with a 'Links' sheet that contains multiple hyperlinks.
Cell C3 has a hyperlink with display text 'View Reports' pointing to http://old.example.com
"""

import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_cop_hyperlink_005'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Links ---
    ws = wb.active
    ws.title = 'Links'

    # Header row
    ws['A1'] = 'Resource Name'
    ws['B1'] = 'Category'
    ws['C1'] = 'Link'
    ws['D1'] = 'Last Updated'
    ws['E1'] = 'Notes'

    for col_letter in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col_letter}1'].font = Font(bold=True)

    # Row 2: A regular hyperlink (not C3)
    ws['A2'] = 'Company Homepage'
    ws['B2'] = 'General'
    ws['C2'] = 'Visit Site'
    ws['C2'].hyperlink = 'https://www.company.com'
    ws['C2'].font = Font(color='0563C1', underline='single')
    ws['D2'] = '2025-01-10'
    ws['E2'] = 'Main website'

    # Row 3: The target hyperlink — C3 with 'View Reports' pointing to OLD URL
    ws['A3'] = 'Monthly Reports'
    ws['B3'] = 'Finance'
    ws['C3'] = 'View Reports'
    ws['C3'].hyperlink = 'http://old.example.com'
    ws['C3'].font = Font(color='0563C1', underline='single')
    ws['D3'] = '2025-02-15'
    ws['E3'] = 'Finance reporting portal'

    # Row 4: Another hyperlink
    ws['A4'] = 'Employee Portal'
    ws['B4'] = 'HR'
    ws['C4'] = 'HR Portal'
    ws['C4'].hyperlink = 'https://hr.internal.company.com/portal'
    ws['C4'].font = Font(color='0563C1', underline='single')
    ws['D4'] = '2025-01-22'
    ws['E4'] = 'Human Resources self-service'

    # Row 5: Another hyperlink
    ws['A5'] = 'Project Tracker'
    ws['B5'] = 'Engineering'
    ws['C5'] = 'Open Tracker'
    ws['C5'].hyperlink = 'https://projects.company.com/tracker'
    ws['C5'].font = Font(color='0563C1', underline='single')
    ws['D5'] = '2025-03-01'
    ws['E5'] = 'Sprint and project management'

    # Row 6: Another hyperlink
    ws['A6'] = 'Knowledge Base'
    ws['B6'] = 'Support'
    ws['C6'] = 'View KB'
    ws['C6'].hyperlink = 'https://kb.company.com'
    ws['C6'].font = Font(color='0563C1', underline='single')
    ws['D6'] = '2025-02-28'
    ws['E6'] = 'Internal documentation wiki'

    # Row 7: Another hyperlink
    ws['A7'] = 'IT Help Desk'
    ws['B7'] = 'IT'
    ws['C7'] = 'Submit Ticket'
    ws['C7'].hyperlink = 'https://helpdesk.company.com/new'
    ws['C7'].font = Font(color='0563C1', underline='single')
    ws['D7'] = '2025-01-30'
    ws['E7'] = 'IT support request system'

    # Row 8: Another hyperlink
    ws['A8'] = 'Sales Dashboard'
    ws['B8'] = 'Sales'
    ws['C8'] = 'View Dashboard'
    ws['C8'].hyperlink = 'https://sales.company.com/dashboard'
    ws['C8'].font = Font(color='0563C1', underline='single')
    ws['D8'] = '2025-03-05'
    ws['E8'] = 'Real-time sales analytics'

    # Row 9: Another hyperlink
    ws['A9'] = 'Cloud Storage'
    ws['B9'] = 'IT'
    ws['C9'] = 'Open Drive'
    ws['C9'].hyperlink = 'https://drive.company.com'
    ws['C9'].font = Font(color='0563C1', underline='single')
    ws['D9'] = '2025-02-05'
    ws['E9'] = 'Shared file storage'

    # Row 10: Another hyperlink
    ws['A10'] = 'Training Platform'
    ws['B10'] = 'HR'
    ws['C10'] = 'Start Learning'
    ws['C10'].hyperlink = 'https://learn.company.com'
    ws['C10'].font = Font(color='0563C1', underline='single')
    ws['D10'] = '2025-02-18'
    ws['E10'] = 'Employee training portal'

    # Row 11: Another hyperlink
    ws['A11'] = 'Expense Reports'
    ws['B11'] = 'Finance'
    ws['C11'] = 'Submit Expenses'
    ws['C11'].hyperlink = 'https://expenses.company.com/submit'
    ws['C11'].font = Font(color='0563C1', underline='single')
    ws['D11'] = '2025-03-10'
    ws['E11'] = 'Employee expense submission'

    # Row 12: Another hyperlink
    ws['A12'] = 'Compliance Documents'
    ws['B12'] = 'Legal'
    ws['C12'] = 'View Policies'
    ws['C12'].hyperlink = 'https://compliance.company.com/policies'
    ws['C12'].font = Font(color='0563C1', underline='single')
    ws['D12'] = '2025-01-05'
    ws['E12'] = 'Legal and compliance resources'

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30

    # --- Sheet 2: Index ---
    ws2 = wb.create_sheet('Index')
    ws2['A1'] = 'Sheet'
    ws2['B1'] = 'Description'
    ws2['A1'].font = Font(bold=True)
    ws2['B1'].font = Font(bold=True)
    ws2['A2'] = 'Links'
    ws2['B2'] = 'Resource links directory'
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 35

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('  Sheet: Links (12 rows, C3 has hyperlink to http://old.example.com with text "View Reports")')


create_initial()
