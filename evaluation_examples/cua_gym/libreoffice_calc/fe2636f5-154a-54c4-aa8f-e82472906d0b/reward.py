"""
Reward Script: Edit hyperlink in cell C3 from old URL to new URL
Task ID: calc_cop_hyperlink_005
Domain: libreoffice_calc
Scoring:
  Component 1: C3 hyperlink URL updated to https://new.example.com/reports  (0.5 pts)
  Component 2: C3 display text remains 'View Reports'                        (0.3 pts)
  Component 3: Other hyperlinks in column C are unchanged                    (0.2 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_cop_hyperlink_005'

# Expected values derived from task description
OLD_URL = 'http://old.example.com'
NEW_URL = 'https://new.example.com/reports'
DISPLAY_TEXT = 'View Reports'

# Other hyperlinks that must not be modified
OTHER_HYPERLINKS = {
    'C2': 'https://www.company.com',
    'C4': 'https://hr.internal.company.com/portal',
    'C5': 'https://projects.company.com/tracker',
    'C6': 'https://kb.company.com',
    'C7': 'https://helpdesk.company.com/new',
    'C8': 'https://sales.company.com/dashboard',
    'C9': 'https://drive.company.com',
    'C10': 'https://learn.company.com',
    'C11': 'https://expenses.company.com/submit',
    'C12': 'https://compliance.company.com/policies',
}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Task: Edit the hyperlink in cell C3 to change its destination URL
    from http://old.example.com to https://new.example.com/reports,
    keeping the display text 'View Reports' unchanged.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook — gate on file validity
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Gate: Sheet 'Links' must exist
    if 'Links' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Links' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Links']

    # Component 1: C3 hyperlink URL updated to https://new.example.com/reports (0.5 points)
    # This is the core task change — the hyperlink target must be updated.
    # On the initial file C3 points to http://old.example.com (FAILS initial)
    # On the golden file C3 points to https://new.example.com/reports (PASSES golden)
    try:
        c3_cell = ws['C3']
        c3_hyperlink = c3_cell.hyperlink
        if c3_hyperlink is None:
            print(f"FAIL: Component 1 — C3 has no hyperlink (expected {NEW_URL})")
        else:
            actual_target = c3_hyperlink.target
            if actual_target == NEW_URL:
                print(f"PASS: Component 1 — C3 hyperlink URL correctly updated to {NEW_URL} (0.5 pts)")
                total_score += 0.5
            elif actual_target == OLD_URL:
                print(f"FAIL: Component 1 — C3 hyperlink URL still points to old URL {OLD_URL}, not yet updated")
            else:
                print(f"FAIL: Component 1 — C3 hyperlink URL is {repr(actual_target)}, expected {NEW_URL}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: C3 display text remains 'View Reports' (0.3 points)
    # The cell value (display text) must remain unchanged throughout the edit.
    # On the initial file, C3.value == 'View Reports' — so this check would pass on initial.
    # BUT we make it conditional: only award points if the hyperlink was already updated AND
    # the display text is preserved. This makes it strictly a task-change verification:
    # the display text MUST be 'View Reports' AND the URL must be the new one.
    try:
        c3_cell = ws['C3']
        c3_hyperlink = c3_cell.hyperlink
        actual_value = c3_cell.value
        actual_target = c3_hyperlink.target if c3_hyperlink else None

        # Only award if both URL is updated AND display text is preserved
        if actual_target == NEW_URL and actual_value == DISPLAY_TEXT:
            print(f"PASS: Component 2 — C3 display text '{DISPLAY_TEXT}' preserved with updated URL (0.3 pts)")
            total_score += 0.3
        elif actual_target != NEW_URL:
            # URL not yet updated, so this is not counted
            print(f"FAIL: Component 2 — URL not updated, cannot verify preservation (display text: {repr(actual_value)})")
        elif actual_value != DISPLAY_TEXT:
            print(f"FAIL: Component 2 — C3 display text changed to {repr(actual_value)}, expected {repr(DISPLAY_TEXT)}")
        else:
            print(f"FAIL: Component 2 — unexpected state")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Other hyperlinks in column C remain unchanged (0.2 points)
    # The task explicitly states no other hyperlinks or cell values should be changed.
    # On the initial file, all these hyperlinks are correct — but this check is conditional:
    # we only award this component if the URL in C3 was updated (i.e., the task was actually acted on).
    # Without this condition, checking other hyperlinks would pass on the initial file too.
    try:
        c3_cell = ws['C3']
        c3_hyperlink = c3_cell.hyperlink
        c3_target = c3_hyperlink.target if c3_hyperlink else None

        if c3_target != NEW_URL:
            # Task not completed — don't award points for preservation
            print("FAIL: Component 3 — skipped (C3 URL not updated, task not completed)")
        else:
            # Task was attempted; verify other hyperlinks are intact
            all_other_intact = True
            for coord, expected_url in OTHER_HYPERLINKS.items():
                cell = ws[coord]
                hl = cell.hyperlink
                if hl is None:
                    print(f"FAIL: Component 3 — {coord} has no hyperlink (expected {expected_url})")
                    all_other_intact = False
                elif hl.target != expected_url:
                    print(f"FAIL: Component 3 — {coord} hyperlink changed to {hl.target} (expected {expected_url})")
                    all_other_intact = False

            if all_other_intact:
                print(f"PASS: Component 3 — All other hyperlinks unchanged (0.2 pts)")
                total_score += 0.2
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
