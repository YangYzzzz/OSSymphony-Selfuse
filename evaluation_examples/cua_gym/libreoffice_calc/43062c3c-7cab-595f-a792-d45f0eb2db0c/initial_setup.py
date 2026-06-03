"""
Initial Setup: financial_report.xlsx with some blank monthly values on Desktop
Task ID: osworld_multi_apps_calc_vscode_011
Domain: libreoffice_calc (multi-app with vscode)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user/Desktop'
TASK_ID = 'osworld_multi_apps_calc_vscode_011'
OUTPUT = f'{WORKDIR}/financial_report.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    os.makedirs(WORKDIR, exist_ok=True)

    wb = openpyxl.Workbook()

    # ---- Sheet 1: Monthly Data ----
    ws = wb.active
    ws.title = "Monthly Data"

    # Header row
    headers = ["Month", "Revenue", "Expenses", "Net Profit", "Units Sold", "Marketing Spend"]
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 2024 monthly data (Jan-Dec) — some values intentionally left blank (None)
    data_2024 = [
        # Month,      Revenue,    Expenses,  Net Profit, Units Sold, Marketing Spend
        ("Jan 2024",  142500.00,  98300.00,  44200.00,   1820,       12400.00),
        ("Feb 2024",  135800.00,  95600.00,  40200.00,   1740,       11900.00),
        ("Mar 2024",  158200.00, 105400.00,  52800.00,   2010,       13500.00),
        ("Apr 2024",  None,       107800.00, None,        2150,       14200.00),  # Revenue/NetProfit blank
        ("May 2024",  175300.00, 109200.00,  66100.00,   2240,       14800.00),
        ("Jun 2024",  182400.00,  None,       None,       2380,       15600.00),  # Expenses/NetProfit blank
        ("Jul 2024",  179600.00, 112300.00,  67300.00,   2290,       15100.00),
        ("Aug 2024",  188500.00, 115700.00,  72800.00,   2410,       15900.00),
        ("Sep 2024",  None,       118200.00, None,        2510,       16300.00),  # Revenue/NetProfit blank
        ("Oct 2024",  195200.00, 120500.00,  74700.00,   2560,       16700.00),
        ("Nov 2024",  210800.00, 128400.00,  82400.00,   None,       17500.00),  # Units Sold blank
        ("Dec 2024",  235600.00, 138900.00,  96700.00,   3020,       19200.00),
    ]

    # 2025 monthly data — some values intentionally left blank
    data_2025 = [
        # Month,      Revenue,    Expenses,  Net Profit, Units Sold, Marketing Spend
        ("Jan 2025",  158400.00, 106200.00,  52200.00,   2020,       13800.00),
        ("Feb 2025",  None,      103400.00,  None,        1940,       13200.00),  # Revenue/NetProfit blank
        ("Mar 2025",  178600.00, 113800.00,  64800.00,   2240,       14900.00),
        ("Apr 2025",  186200.00, 117600.00,  68600.00,   2390,       15700.00),
        ("May 2025",  None,       118900.00, None,        2490,       16400.00),  # Revenue/NetProfit blank
        ("Jun 2025",  204800.00, 124600.00,  80200.00,   2650,       17200.00),
        ("Jul 2025",  198400.00, 122100.00,  76300.00,   2540,       16800.00),
        ("Aug 2025",  212600.00, 128500.00,  84100.00,   2720,       17800.00),
        ("Sep 2025",  None,       134800.00, None,        2840,       18600.00),  # Revenue/NetProfit blank
        ("Oct 2025",  228400.00, 138900.00,  89500.00,   2920,       19100.00),
        ("Nov 2025",  245600.00, 148200.00,  97400.00,   None,       20300.00),  # Units Sold blank
        ("Dec 2025",  268900.00, 159400.00, 109500.00,   3460,       22100.00),
    ]

    all_data = data_2024 + data_2025
    for r, row_data in enumerate(all_data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 18

    ws.freeze_panes = "A2"

    # ---- Sheet 2: Notes ----
    ws2 = wb.create_sheet("Notes")
    ws2["A1"] = "Data Notes"
    ws2["A1"].font = Font(bold=True, size=13)
    ws2["A3"] = "Source:"
    ws2["B3"] = "Internal Finance Department"
    ws2["A4"] = "Updated:"
    ws2["B4"] = "2025-12-31"
    ws2["A5"] = "Currency:"
    ws2["B5"] = "USD"
    ws2["A7"] = "Note:"
    ws2["B7"] = "Some monthly values are missing due to data collection issues."
    ws2["B7"].alignment = Alignment(wrap_text=True)
    ws2["A8"] = "Action:"
    ws2["B8"] = "Export to CSV and interpolate missing values using linear interpolation before analysis."
    ws2["B8"].alignment = Alignment(wrap_text=True)
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 60

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the xlsx in LibreOffice Calc and VSCode
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.5)
    launch_gui('code /home/user/Desktop', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc and VSCode with DISPLAY=:0')


create_initial()
