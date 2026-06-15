"""
Reward Script: Create a multi-ring doughnut chart comparing this year's and last year's budget allocation
Task ID: calc_chart_doughnut_multi_070
Domain: libreoffice_calc
Scoring:
  Component 1: DoughnutChart exists on 'BudgetCompare' sheet (0.3 pts)
  Component 2: Chart title is 'Budget Allocation: This Year vs Last Year' (0.3 pts)
  Component 3: Chart has 2 series with correct data references covering both columns (0.4 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_chart_doughnut_multi_070'
SHEET_NAME = 'BudgetCompare'
EXPECTED_TITLE = 'Budget Allocation: This Year vs Last Year'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: BudgetCompare sheet must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Available sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Component 1: A DoughnutChart exists on the BudgetCompare sheet (0.3 points)
    # Initial file has 0 charts; golden file has 1 DoughnutChart. Only earns points when chart exists.
    try:
        charts = ws._charts
        doughnut_charts = [c for c in charts if type(c).__name__ == 'DoughnutChart']
        if len(doughnut_charts) >= 1:
            print(f"PASS: Component 1 — DoughnutChart exists on '{SHEET_NAME}' (found {len(doughnut_charts)}) (0.3 pts)")
            total_score += 0.3
            target_chart = doughnut_charts[0]
        else:
            all_types = [type(c).__name__ for c in charts]
            print(f"FAIL: Component 1 — Expected at least 1 DoughnutChart on '{SHEET_NAME}', found: {all_types}")
            target_chart = None
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        target_chart = None

    # Component 2: Chart title is exactly 'Budget Allocation: This Year vs Last Year' (0.3 points)
    # This check requires a chart to exist (Component 1 must pass first).
    if target_chart is not None:
        try:
            # Navigate the title structure: chart.title.tx.rich.p[0].r[0].t
            title_text = None
            if target_chart.title is not None:
                try:
                    title_text = target_chart.title.tx.rich.p[0].r[0].t
                except Exception:
                    pass
                # Fallback: try tx.rich.p[0] paragraphs
                if title_text is None:
                    try:
                        for para in target_chart.title.tx.rich.p:
                            for run in para.r:
                                if run.t:
                                    title_text = run.t
                                    break
                            if title_text:
                                break
                    except Exception:
                        pass

            if title_text and title_text.strip() == EXPECTED_TITLE:
                print(f"PASS: Component 2 — Chart title is '{title_text}' (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — Expected title '{EXPECTED_TITLE}', found: '{title_text}'")
        except Exception as e:
            print(f"ERROR: Component 2 — {e}")
    else:
        print("SKIP: Component 2 — No DoughnutChart to check title")

    # Component 3: Chart has 2 series covering both 'This Year' (col B) and 'Last Year' (col C) data (0.4 points)
    # Multi-ring doughnut requires exactly 2 series, each referencing the correct data columns.
    # This FAILS on initial (no chart) and PASSES on golden (2 series with B2:B5 and C2:C5).
    if target_chart is not None:
        try:
            series = target_chart.series
            if len(series) >= 2:
                # Check series 0 references col B (This Year) and series 1 references col C (Last Year)
                # OR either order is acceptable since the task just requires both columns to be represented.
                refs_found = []
                for ser in series:
                    try:
                        if ser.val and ser.val.numRef:
                            refs_found.append(ser.val.numRef.f)
                    except Exception:
                        pass

                # Verify that both column B data and column C data are represented
                has_col_b = any('$B$' in ref or "'BudgetCompare'!$B" in ref or "BudgetCompare'!B" in ref or "'BudgetCompare'!B" in ref for ref in refs_found)
                has_col_c = any('$C$' in ref or "'BudgetCompare'!$C" in ref or "BudgetCompare'!C" in ref or "'BudgetCompare'!C" in ref for ref in refs_found)

                if has_col_b and has_col_c:
                    print(f"PASS: Component 3 — Chart has {len(series)} series covering both This Year (col B) and Last Year (col C) data (0.4 pts)")
                    print(f"  Series refs: {refs_found}")
                    total_score += 0.4
                elif len(series) >= 2:
                    # 2 series exist but refs don't match expected columns — partial match
                    # Award points if we have 2 series (at least the two-ring structure is present)
                    print(f"PARTIAL: Component 3 — Chart has {len(series)} series but column refs unexpected: {refs_found}")
                    print(f"  Expected refs covering col B and col C. has_col_b={has_col_b}, has_col_c={has_col_c}")
                    total_score += 0.2
                else:
                    print(f"FAIL: Component 3 — Chart has {len(series)} series (need 2 for multi-ring), refs: {refs_found}")
            else:
                print(f"FAIL: Component 3 — Chart has only {len(series)} series (need at least 2 for multi-ring doughnut)")
        except Exception as e:
            print(f"ERROR: Component 3 — {e}")
    else:
        print("SKIP: Component 3 — No DoughnutChart to check series")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
