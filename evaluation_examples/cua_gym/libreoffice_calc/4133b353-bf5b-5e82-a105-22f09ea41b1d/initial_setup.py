"""
Initial Setup: Create spreadsheet with budget comparison data (no charts)
Task ID: calc_chart_doughnut_multi_070
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_doughnut_multi_070'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: BudgetCompare ---
    ws = wb.active
    ws.title = 'BudgetCompare'

    # Headers
    headers = ['Department', 'This Year', 'Last Year']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Data rows (exact values from task context)
    data = [
        ['Engineering', 480000, 420000],
        ['Sales', 320000, 310000],
        ['Marketing', 210000, 180000],
        ['Support', 150000, 140000],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

    # No charts in initial file
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: BudgetCompare')
    print('Data rows: 4 departments with This Year and Last Year budget values')
    print('No charts present (task: agent must create the doughnut chart)')


create_initial()
