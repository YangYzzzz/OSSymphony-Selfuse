"""
Reward Script: Organize volunteer schedule for school fundraising event
Task ID: calc_edu_volunteer_schedule_022
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Column F has COUNTA(C#:E#) formulas for rows 2-41              — 0.30 pts
  Component 2: B44:B48 have COUNTIF formulas counting each time slot           — 0.25 pts
  Component 3: C44:C48 have IF formulas flagging understaffed slots (<4)       — 0.20 pts
  Component 4: Data rows A2:A41 sorted ascending by last name                  — 0.15 pts
  Component 5: Conditional formatting on B44:B48 (red fill when < 4)          — 0.10 pts
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_volunteer_schedule_022'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Volunteers' not in wb.sheetnames:
        print("CRITICAL: 'Volunteers' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Volunteers']

    # Component 1: Column F has COUNTA(C#:E#) formulas for all 40 data rows (0.30 pts)
    # Task: "Count how many volunteers are assigned to each time slot" via Total Shifts column
    # This FAILS on initial (all None) -> PASSES on golden (all have COUNTA formulas)
    try:
        counta_count = 0
        counta_pattern = re.compile(r'=COUNTA\(C\d+:E\d+\)', re.IGNORECASE)
        for row in range(2, 42):
            cell_val = ws.cell(row=row, column=6).value
            if cell_val and isinstance(cell_val, str) and counta_pattern.match(cell_val.replace(' ', '')):
                counta_count += 1

        if counta_count == 40:
            print(f"PASS: Component 1 — All 40 rows in column F have COUNTA formulas ({counta_count}/40) (0.30 pts)")
            total_score += 0.30
        elif counta_count >= 30:
            print(f"PARTIAL: Component 1 — {counta_count}/40 rows in column F have COUNTA formulas (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — Only {counta_count}/40 rows in column F have COUNTA formulas (expected 40)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: B44:B48 have COUNTIF formulas counting each time slot (0.25 pts)
    # Task: "Count how many volunteers are assigned to each time slot"
    # FAILS on initial (all None) -> PASSES on golden (all have COUNTIF formulas)
    try:
        countif_count = 0
        countif_pattern = re.compile(r'=COUNTIF\(', re.IGNORECASE)
        for row in range(44, 49):
            cell_val = ws.cell(row=row, column=2).value
            if cell_val and isinstance(cell_val, str) and countif_pattern.search(cell_val):
                countif_count += 1

        if countif_count == 5:
            print(f"PASS: Component 2 — All 5 COUNTIF formulas present in B44:B48 ({countif_count}/5) (0.25 pts)")
            total_score += 0.25
        elif countif_count >= 3:
            print(f"PARTIAL: Component 2 — {countif_count}/5 COUNTIF formulas in B44:B48 (0.12 pts)")
            total_score += 0.12
        else:
            print(f"FAIL: Component 2 — Only {countif_count}/5 COUNTIF formulas in B44:B48 (expected 5)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: C44:C48 have IF formulas flagging understaffed slots (<4) (0.20 pts)
    # Task: "flag any slots that are understaffed (fewer than 4 volunteers)"
    # FAILS on initial (all None) -> PASSES on golden (all have IF formulas)
    try:
        if_count = 0
        # Pattern: =IF(B##<4,"Yes","No")
        if_pattern = re.compile(r'=IF\(B\d+<4,', re.IGNORECASE)
        for row in range(44, 49):
            cell_val = ws.cell(row=row, column=3).value
            if cell_val and isinstance(cell_val, str) and if_pattern.search(cell_val.replace(' ', '')):
                if_count += 1

        if if_count == 5:
            print(f"PASS: Component 3 — All 5 IF(<4) formulas present in C44:C48 ({if_count}/5) (0.20 pts)")
            total_score += 0.20
        elif if_count >= 3:
            print(f"PARTIAL: Component 3 — {if_count}/5 IF formulas in C44:C48 (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — Only {if_count}/5 IF(<4) formulas in C44:C48 (expected 5)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Data sorted ascending by last name (A2:A41) (0.15 pts)
    # Task: "sort the volunteer list by last name"
    # FAILS on initial (unsorted) -> PASSES on golden (sorted A-Z)
    try:
        last_names = []
        for row in range(2, 42):
            val = ws.cell(row=row, column=1).value
            if val is not None:
                last_names.append(str(val))

        if len(last_names) == 40 and last_names == sorted(last_names):
            print(f"PASS: Component 4 — 40 last names sorted ascending A-Z (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — Last names not sorted. First 5: {last_names[:5]}, Expected sorted: {sorted(last_names[:5])}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Conditional formatting on B44:B48 with red fill when value < 4 (0.10 pts)
    # Task: "Conditional formatting on B44:B48: red if count < 4"
    # FAILS on initial (no CF rules) -> PASSES on golden (CF rule with red fill, <4)
    try:
        cf_rules = ws.conditional_formatting._cf_rules
        cf_found = False
        for cf_range, rules in cf_rules.items():
            cf_range_str = str(cf_range)
            if 'B44' in cf_range_str and 'B48' in cf_range_str:
                for rule in rules:
                    # Check: cellIs type, lessThan operator, formula ['4'], red fill
                    if (rule.type == 'cellIs' and
                            rule.operator == 'lessThan' and
                            rule.formula == ['4'] and
                            rule.dxf and rule.dxf.fill and
                            rule.dxf.fill.fill_type == 'solid'):
                        # Check for red color (FFFF0000 or variants with red channel)
                        try:
                            fg_rgb = rule.dxf.fill.fgColor.rgb
                            # Red: FFFF0000, or at least high red channel
                            if fg_rgb and 'FF0000' in fg_rgb.upper():
                                cf_found = True
                                print(f"PASS: Component 5 — CF rule on B44:B48: cellIs<4, red fill ({fg_rgb}) (0.10 pts)")
                        except Exception:
                            pass

        if cf_found:
            total_score += 0.10
        else:
            # Check if CF range exists even without perfect color match
            has_cf_range = any('B44' in str(r) and 'B48' in str(r) for r in cf_rules.keys())
            if has_cf_range:
                # Partial: CF range exists but color doesn't match exactly
                for cf_range, rules in cf_rules.items():
                    cf_range_str = str(cf_range)
                    if 'B44' in cf_range_str and 'B48' in cf_range_str:
                        for rule in rules:
                            if rule.type == 'cellIs' and rule.operator == 'lessThan':
                                print(f"PARTIAL: Component 5 — CF rule on B44:B48 exists (cellIs<{rule.formula}) but fill color check failed (0.05 pts)")
                                total_score += 0.05
                                cf_found = True
                                break
                if not cf_found:
                    print(f"FAIL: Component 5 — CF range B44:B48 found but rule type/operator mismatch")
            else:
                print(f"FAIL: Component 5 — No conditional formatting found on B44:B48")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
