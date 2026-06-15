"""
Initial Setup: Volunteer Schedule for School Fundraising Event
Task ID: calc_edu_volunteer_schedule_022
Domain: libreoffice_calc

Creates a spreadsheet with 40 volunteers each assigned up to 3 time slots.
The Volunteers sheet has a summary section below the data but without formulas
or sorted order — those are added as part of the task.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_volunteer_schedule_022'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Volunteers'

    # --- Headers ---
    headers = ['Last Name', 'First Name', 'Slot1', 'Slot2', 'Slot3', 'Total Shifts']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')

    # Time slots available
    slots = ['Morning Setup', 'Afternoon Booth', 'Evening Cleanup']

    # 40 volunteers (realistic names, NOT sorted — task will sort them)
    volunteers = [
        ('Ramirez',    'Sofia',    'Afternoon Booth', 'Evening Cleanup', ''),
        ('Nguyen',     'David',    'Morning Setup',   'Afternoon Booth', 'Evening Cleanup'),
        ('Williams',   'Jasmine',  'Morning Setup',   '',                ''),
        ('Patel',      'Arjun',    'Afternoon Booth', '',                ''),
        ('Johnson',    'Marcus',   'Evening Cleanup', 'Morning Setup',   ''),
        ('Chen',       'Lily',     'Morning Setup',   'Afternoon Booth', ''),
        ('Thompson',   'Brandon',  'Evening Cleanup', '',                ''),
        ('Garcia',     'Isabella', 'Morning Setup',   'Evening Cleanup', 'Afternoon Booth'),
        ('Martinez',   'Carlos',   'Afternoon Booth', 'Morning Setup',   ''),
        ('Anderson',   'Megan',    'Evening Cleanup', 'Afternoon Booth', ''),
        ('Taylor',     'Jordan',   'Morning Setup',   '',                ''),
        ('Harris',     'Aisha',    'Afternoon Booth', 'Evening Cleanup', ''),
        ('Wilson',     'Ethan',    'Evening Cleanup', 'Morning Setup',   'Afternoon Booth'),
        ('Brown',      'Priya',    'Morning Setup',   'Afternoon Booth', ''),
        ('Davis',      'Chloe',    'Afternoon Booth', '',                ''),
        ('Miller',     'Tyler',    'Evening Cleanup', 'Afternoon Booth', ''),
        ('Moore',      'Samantha', 'Morning Setup',   'Evening Cleanup', ''),
        ('Jackson',    'Darius',   'Afternoon Booth', 'Morning Setup',   'Evening Cleanup'),
        ('White',      'Natalie',  'Morning Setup',   '',                ''),
        ('Lopez',      'Diego',    'Evening Cleanup', '',                ''),
        ('Clark',      'Hailey',   'Afternoon Booth', 'Morning Setup',   ''),
        ('Lewis',      'Noah',     'Morning Setup',   'Afternoon Booth', 'Evening Cleanup'),
        ('Robinson',   'Kayla',    'Evening Cleanup', 'Afternoon Booth', ''),
        ('Walker',     'Elijah',   'Morning Setup',   '',                ''),
        ('Hall',       'Emma',     'Afternoon Booth', 'Evening Cleanup', ''),
        ('Young',      'Liam',     'Evening Cleanup', 'Morning Setup',   ''),
        ('Allen',      'Grace',    'Morning Setup',   'Afternoon Booth', ''),
        ('King',       'James',    'Afternoon Booth', '',                ''),
        ('Wright',     'Olivia',   'Evening Cleanup', 'Morning Setup',   'Afternoon Booth'),
        ('Scott',      'Lucas',    'Morning Setup',   'Evening Cleanup', ''),
        ('Torres',     'Maria',    'Afternoon Booth', 'Morning Setup',   ''),
        ('Hernandez',  'Elena',    'Morning Setup',   '',                ''),
        ('Green',      'Xavier',   'Evening Cleanup', 'Afternoon Booth', ''),
        ('Adams',      'Brianna',  'Afternoon Booth', 'Evening Cleanup', 'Morning Setup'),
        ('Nelson',     'Kevin',    'Morning Setup',   'Afternoon Booth', ''),
        ('Baker',      'Alexis',   'Evening Cleanup', '',                ''),
        ('Carter',     'Isaiah',   'Morning Setup',   'Evening Cleanup', ''),
        ('Mitchell',   'Vanessa',  'Afternoon Booth', 'Morning Setup',   'Evening Cleanup'),
        ('Perez',      'Adrian',   'Evening Cleanup', 'Afternoon Booth', ''),
        ('Roberts',    'Faith',    'Morning Setup',   '',                ''),
    ]

    # Write volunteer data rows 2-41 (NOT sorted — task will sort)
    for r, (last, first, s1, s2, s3) in enumerate(volunteers, 2):
        ws.cell(row=r, column=1, value=last)
        ws.cell(row=r, column=2, value=first)
        ws.cell(row=r, column=3, value=s1 if s1 else None)
        ws.cell(row=r, column=4, value=s2 if s2 else None)
        ws.cell(row=r, column=5, value=s3 if s3 else None)
        # Column F (Total Shifts) intentionally left empty — task adds COUNTA formula

    # --- Summary section (rows 43-48) ---
    # Row 43: Headers for summary table
    ws.cell(row=43, column=1, value='Time Slot').font = Font(bold=True)
    ws.cell(row=43, column=2, value='Volunteer Count').font = Font(bold=True)
    ws.cell(row=43, column=3, value='Understaffed').font = Font(bold=True)

    # Rows 44-48: Slot names — counts and understaffed flags left empty (task adds formulas)
    time_slot_names = [
        'Morning Setup',
        'Afternoon Booth',
        'Evening Cleanup',
        'Registration Desk',
        'Raffle Table',
    ]
    for i, slot_name in enumerate(time_slot_names):
        row = 44 + i
        ws.cell(row=row, column=1, value=slot_name)
        # B and C columns intentionally left empty — task adds COUNTIF and IF formulas

    # Column widths for readability
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('  - Sheet: Volunteers')
    print('  - Rows 2-41: 40 volunteers with time slots (unsorted)')
    print('  - Rows 43-48: Summary section headers and slot names only')
    print('  - Column F (Total Shifts): empty')
    print('  - B44:B48 (Volunteer Count): empty')
    print('  - C44:C48 (Understaffed): empty')


create_initial()
