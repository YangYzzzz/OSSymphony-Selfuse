"""
Initial Setup: Salesperson commission spreadsheet with partial formula in D2 only.
Task ID: osworld_calc_formula_pattern_concat_011
Domain: libreoffice_calc

Creates a spreadsheet with:
- Column A: Name
- Column B: Sales Amount
- Column C: Commission Rate
- Column D: Commission (formula =B2*C2 in D2 only; D3:D11 are empty)
- Column E: empty (task is to fill with concatenation formulas)
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_formula_pattern_concat_011'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Sales ---
    ws = wb.active
    ws.title = 'Sales'

    # Headers
    headers = ['Name', 'Sales Amount', 'Commission Rate', 'Commission']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic salesperson data (10 rows)
    data = [
        ['Sarah Chen',       92450.00, 0.08],
        ['Marcus Johnson',   78200.50, 0.07],
        ['Emily Rodriguez',  115300.75, 0.09],
        ['Daniel Kim',       63800.00, 0.06],
        ['Rachel Thompson',  88600.25, 0.08],
        ['James O\'Brien',   54200.00, 0.05],
        ['Priya Patel',      101750.50, 0.09],
        ['Aaron Williams',   47300.00, 0.05],
        ['Natalie Foster',   136500.00, 0.10],
        ['Kevin Martinez',   72400.75, 0.07],
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])  # Name
        ws.cell(row=r, column=2, value=row_data[1])  # Sales Amount
        ws.cell(row=r, column=3, value=row_data[2])  # Commission Rate
        # Column D: only D2 has a formula; D3:D11 are left empty

    # D2 has the starter formula only
    ws.cell(row=2, column=4, value='=B2*C2')

    # Column E is intentionally left empty — task is to fill it

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
