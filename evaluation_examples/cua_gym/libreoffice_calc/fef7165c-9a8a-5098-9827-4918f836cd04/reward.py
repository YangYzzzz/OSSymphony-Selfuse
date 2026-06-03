"""
Reward Script: Extend commission formula down column D and create concatenation formulas in column E
Task ID: osworld_calc_formula_pattern_concat_011
Domain: libreoffice_calc
Scoring:
  - Component 1: Commission formulas filled in D3:D11 (0.4 pts)
  - Component 2: Column E header 'Description' exists in E1 (0.1 pts)
  - Component 3: Concatenation formulas in E2:E11 with correct pattern (0.5 pts)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_formula_pattern_concat_011'

# Number of data rows (rows 2 through 11 = 10 rows of salesperson data)
DATA_ROWS = list(range(2, 12))  # rows 2–11


def normalize_formula(formula):
    """Normalize formula string for comparison: uppercase, strip spaces."""
    if not isinstance(formula, str):
        return ""
    return formula.upper().replace(" ", "")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the Sales sheet exists
    if 'Sales' not in wb.sheetnames:
        print("CRITICAL: 'Sales' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Sales']

    # ---------------------------------------------------------------------------
    # Component 1: Commission formulas extended to D3:D11 (0.4 pts)
    # In initial_env D2 has =B2*C2 but D3:D11 are empty (None).
    # The task requires filling D3:D11 with corresponding =Bn*Cn formulas.
    # ---------------------------------------------------------------------------
    try:
        d_formula_count = 0
        d_expected_total = len(DATA_ROWS) - 1  # rows 3–11 = 9 rows to extend
        for row in DATA_ROWS[1:]:  # skip row 2 which already existed
            cell_val = ws.cell(row=row, column=4).value  # column D
            if cell_val is not None:
                # Accept any formula or numeric value referencing the row
                norm = normalize_formula(str(cell_val))
                # Expect something like =B3*C3 pattern: =B{row}*C{row}
                expected_pattern = f"=B{row}*C{row}"
                if norm == expected_pattern.upper():
                    d_formula_count += 1
                elif isinstance(cell_val, (int, float)) and cell_val > 0:
                    # Might be a hard-coded numeric value (less ideal but task may allow it)
                    d_formula_count += 0.5
                else:
                    print(f"FAIL: D{row} has unexpected value: {repr(cell_val)}")
            else:
                print(f"FAIL: D{row} is empty (expected commission formula)")

        fraction_filled = d_formula_count / d_expected_total
        component1_score = round(0.4 * fraction_filled, 4)
        if fraction_filled >= 1.0:
            print(f"PASS: Component 1 — All D3:D11 commission formulas present ({d_formula_count}/{d_expected_total}) (0.4 pts)")
            total_score += component1_score
        elif fraction_filled > 0:
            print(f"PARTIAL: Component 1 — {d_formula_count}/{d_expected_total} D-column formulas filled ({component1_score:.4f} pts)")
            total_score += component1_score
        else:
            print(f"FAIL: Component 1 — No commission formulas found in D3:D11 (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ---------------------------------------------------------------------------
    # Component 2: Column E header 'Description' in E1 (0.1 pts)
    # This is a new column not present in initial_env (max_col was 4 initially).
    # ---------------------------------------------------------------------------
    try:
        e1_val = ws.cell(row=1, column=5).value  # E1
        if e1_val is not None and str(e1_val).strip().lower() == 'description':
            print(f"PASS: Component 2 — E1 has 'Description' header (0.1 pts)")
            total_score += 0.1
        else:
            print(f"FAIL: Component 2 — E1 expected 'Description', found: {repr(e1_val)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---------------------------------------------------------------------------
    # Component 3: Concatenation formulas in E2:E11 with correct pattern (0.5 pts)
    # Expected pattern per row: =A{r}&" | Sales: $"&TEXT(B{r},"0.00")&" | Rate: "&TEXT(C{r},"0.00")&" | Comm: $"&TEXT(D{r},"0.00")
    # We check each row for:
    #   a) A formula string (starts with '=')
    #   b) Contains &TEXT( pattern
    #   c) References A, B, C, D for that row
    # ---------------------------------------------------------------------------
    try:
        e_correct = 0
        e_total = len(DATA_ROWS)  # rows 2–11

        for row in DATA_ROWS:
            cell_val = ws.cell(row=row, column=5).value  # column E
            if cell_val is None:
                print(f"FAIL: E{row} is empty (expected concatenation formula)")
                continue

            val_str = str(cell_val)
            if not val_str.startswith('='):
                print(f"FAIL: E{row} is not a formula: {repr(cell_val)}")
                continue

            norm = normalize_formula(val_str)

            # Check for core structural elements:
            # 1. References correct row cells: A{row}, B{row}, C{row}, D{row}
            refs_ok = (
                f"A{row}" in norm and
                f"B{row}" in norm and
                f"C{row}" in norm and
                f"D{row}" in norm
            )
            # 2. Uses TEXT() function for formatting
            has_text_func = "TEXT(" in norm
            # 3. Uses & concatenation operator
            has_concat = "&" in norm
            # 4. Contains dollar sign and pipe separators in string literals
            has_dollar = '"$"' in val_str.upper() or '$"' in val_str or '"$' in val_str
            has_pipe = '" | "' in val_str or '| ' in val_str or ' |' in val_str

            if refs_ok and has_text_func and has_concat:
                # Full credit: formula structure is correct
                e_correct += 1
                print(f"PASS: E{row} — valid concatenation formula")
            else:
                issues = []
                if not refs_ok:
                    issues.append(f"missing correct row references (A{row}/B{row}/C{row}/D{row})")
                if not has_text_func:
                    issues.append("missing TEXT() function")
                if not has_concat:
                    issues.append("missing & concatenation")
                print(f"FAIL: E{row} — formula issues: {', '.join(issues)} — value: {repr(cell_val)}")

        fraction_e = e_correct / e_total
        component3_score = round(0.5 * fraction_e, 4)
        if fraction_e >= 1.0:
            print(f"PASS: Component 3 — All E2:E11 concatenation formulas correct ({e_correct}/{e_total}) (0.5 pts)")
            total_score += component3_score
        elif fraction_e > 0:
            print(f"PARTIAL: Component 3 — {e_correct}/{e_total} E-column formulas valid ({component3_score:.4f} pts)")
            total_score += component3_score
        else:
            print(f"FAIL: Component 3 — No valid concatenation formulas in E2:E11 (0.0 pts)")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
