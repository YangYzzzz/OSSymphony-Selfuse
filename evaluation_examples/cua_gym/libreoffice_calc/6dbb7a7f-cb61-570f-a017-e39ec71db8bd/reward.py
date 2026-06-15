"""
Reward Script: Equipment Loan Comparison Table
Task ID: calc_fin_loan_comparison_053
Domain: libreoffice_calc
Scoring:
  - Component 1: PMT formulas in B7:D7 (0.30 pts)
  - Component 2: Derived calculation formulas in B8:D10 (0.20 pts)
  - Component 3: Currency format ($#,##0.00) on B7:D10 (0.15 pts)
  - Component 4: Conditional formatting on B10:D10 (highlight min in green) (0.20 pts)
  - Component 5: Sheet protection enabled (0.10 pts)
  - Component 6: Row 3 headers bold (0.05 pts)
  Total: 1.0
"""

import os
import openpyxl
from openpyxl.utils import get_column_letter

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_loan_comparison_053'


def check_pmt_formula(val, expected_core):
    """Return 1.0 if formula matches, 0.5 if has PMT but wrong args, 0.0 otherwise."""
    if val and isinstance(val, str) and 'PMT' in val.upper():
        val_norm = val.upper().replace(' ', '').replace('=', '')
        exp_norm = expected_core.upper().replace(' ', '')
        if exp_norm in val_norm:
            return 1.0
        return 0.5
    return 0.0


def check_derived_formula(val, expected_core):
    """Return True if cell contains a formula with the expected core expression."""
    if val and isinstance(val, str):
        val_norm = val.upper().replace(' ', '').replace('=', '')
        exp_norm = expected_core.upper().replace(' ', '')
        return exp_norm in val_norm
    return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet 'LoanComparison' exists
    if 'LoanComparison' not in wb.sheetnames:
        print("CRITICAL: Sheet 'LoanComparison' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['LoanComparison']

    # Component 1: PMT formulas in B7, C7, D7 (0.30 points)
    # These cells must contain PMT formula strings referencing the correct cells
    # FAILS on initial (cells are empty), PASSES on golden (PMT formulas present)
    try:
        pmt_checks = {
            'B7': (ws['B7'].value, 'PMT(B5/12,B6*12,-B4)'),
            'C7': (ws['C7'].value, 'PMT(C5/12,C6*12,-C4)'),
            'D7': (ws['D7'].value, 'PMT(D5/12,D6*12,-D4)'),
        }

        pmt_passed = 0.0
        for coord, (val, expected_core) in pmt_checks.items():
            result = check_pmt_formula(val, expected_core)
            if result == 1.0:
                print(f"PASS: {coord} has correct PMT formula: {val}")
                pmt_passed += 1.0
            elif result == 0.5:
                print(f"PARTIAL: {coord} has PMT formula but wrong args: {val} (expected: {expected_core})")
                pmt_passed += 0.5
            else:
                print(f"FAIL: {coord} missing PMT formula, found: {repr(val)}")

        # Award points proportionally based on how many PMT formulas are correct
        pmt_score = (pmt_passed / 3.0) * 0.30
        if pmt_score > 0:
            total_score += pmt_score
        print(f"Component 1 score: {pmt_score:.2f}/0.30 ({pmt_passed}/3 PMT formulas correct)")
    except Exception as e:
        print(f"ERROR: Component 1 (PMT formulas) — {e}")

    # Component 2: Derived calculation formulas (B8:D10) (0.20 points)
    # B8: =B7*B6*12, C8: =C7*C6*12, D8: =D7*D6*12
    # B9: =B8-B4, C9: =C8-C4, D9: =D8-D4
    # B10: =B8, C10: =C8, D10: =D8
    # FAILS on initial (cells empty), PASSES on golden (formulas present)
    try:
        formula_checks = {
            'B8': (ws['B8'].value, 'B7*B6*12'),
            'C8': (ws['C8'].value, 'C7*C6*12'),
            'D8': (ws['D8'].value, 'D7*D6*12'),
            'B9': (ws['B9'].value, 'B8-B4'),
            'C9': (ws['C9'].value, 'C8-C4'),
            'D9': (ws['D9'].value, 'D8-D4'),
            'B10': (ws['B10'].value, 'B8'),
            'C10': (ws['C10'].value, 'C8'),
            'D10': (ws['D10'].value, 'D8'),
        }

        formula_passed = 0
        for coord, (val, expected_core) in formula_checks.items():
            if check_derived_formula(val, expected_core):
                print(f"PASS: {coord} has correct formula: {val}")
                formula_passed += 1
            else:
                print(f"FAIL: {coord} missing/wrong formula: {repr(val)} (expected core: {expected_core})")

        formula_score = (formula_passed / 9.0) * 0.20
        if formula_score > 0:
            total_score += formula_score
        print(f"Component 2 score: {formula_score:.2f}/0.20 ({formula_passed}/9 derived formulas correct)")
    except Exception as e:
        print(f"ERROR: Component 2 (derived formulas) — {e}")

    # Component 3: Currency format on B7:D10 (0.15 points)
    # In initial: 'General' format; in golden: '$#,##0.00'
    # FAILS on initial (General format), PASSES on golden (currency format)
    try:
        currency_count = 0
        total_currency_cells = 12  # 3 cols x 4 rows (B-D, rows 7-10)
        for row in range(7, 11):
            for col in range(2, 5):
                cell = ws.cell(row=row, column=col)
                fmt = cell.number_format or ''
                if '$' in fmt or ('#,##0' in fmt and fmt != 'General'):
                    currency_count += 1

        if currency_count == total_currency_cells:
            print(f"PASS: All {total_currency_cells} cells in B7:D10 have currency format")
            total_score += 0.15
        elif currency_count > 0:
            partial_currency = (currency_count / total_currency_cells) * 0.15
            print(f"PARTIAL: {currency_count}/{total_currency_cells} cells in B7:D10 have currency format")
            total_score += partial_currency
        else:
            print(f"FAIL: No cells in B7:D10 have currency format (sample: B7={ws['B7'].number_format!r})")
    except Exception as e:
        print(f"ERROR: Component 3 (currency format) — {e}")

    # Component 4: Conditional formatting on B10:D10 — highlight minimum in green (0.20 points)
    # Expected: FormulaRule on B10:D10 with MIN formula + green fill
    # FAILS on initial (no CF at all), PASSES on golden (CF with MIN+green)
    try:
        cf_range_found = False
        cf_min_formula_found = False
        cf_green_fill_found = False

        for cf_range, cf_rules in ws.conditional_formatting._cf_rules.items():
            range_str = str(cf_range)
            # Check if the range covers row 10
            if 'B10' in range_str or ('10' in range_str and 'D10' in range_str):
                cf_range_found = True
                for rule in cf_rules:
                    # Check for formula mentioning MIN
                    if hasattr(rule, 'formula') and rule.formula:
                        for f in rule.formula:
                            if 'MIN' in str(f).upper():
                                cf_min_formula_found = True
                                print(f"PASS: Conditional formatting formula references MIN: {f}")
                    # Check for green fill
                    if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                        try:
                            fill_color = rule.dxf.fill.fgColor.rgb
                            if fill_color and len(fill_color) >= 8:
                                r = int(fill_color[2:4], 16)
                                g = int(fill_color[4:6], 16)
                                b = int(fill_color[6:8], 16)
                                if g > 100 and g > r and g > b:
                                    cf_green_fill_found = True
                                    print(f"PASS: Conditional formatting fill is green: {fill_color}")
                                else:
                                    print(f"FAIL: CF fill color not green: {fill_color}")
                        except Exception as color_err:
                            print(f"WARN: Could not read fill color: {color_err}")

        if not cf_range_found:
            print("FAIL: No conditional formatting found covering row 10 (B10:D10)")
        elif cf_min_formula_found and cf_green_fill_found:
            print("PASS: Conditional formatting complete — MIN formula + green fill on B10:D10")
            total_score += 0.20
        elif cf_min_formula_found or cf_green_fill_found:
            print(f"PARTIAL: CF partially correct (min_formula={cf_min_formula_found}, green_fill={cf_green_fill_found})")
            total_score += 0.10
        else:
            print("FAIL: CF found on row 10 but neither MIN formula nor green fill verified")
    except Exception as e:
        print(f"ERROR: Component 4 (conditional formatting) — {e}")

    # Component 5: Sheet protection enabled (0.10 points)
    # In initial: ws.protection.sheet = False; in golden: True
    # FAILS on initial, PASSES on golden
    try:
        if ws.protection.sheet:
            print(f"PASS: Sheet protection is enabled")
            total_score += 0.10
        else:
            print(f"FAIL: Sheet protection is not enabled (ws.protection.sheet = {ws.protection.sheet})")
    except Exception as e:
        print(f"ERROR: Component 5 (sheet protection) — {e}")

    # Component 6: Row 3 headers bold (0.05 points)
    # In initial: row 3 not bold; in golden: row 3 bold
    # FAILS on initial, PASSES on golden
    try:
        bold_count = sum(
            1 for col in range(1, 5)
            if ws.cell(row=3, column=col).font.bold
        )

        if bold_count == 4:
            print(f"PASS: All row 3 header cells (A3:D3) are bold")
            total_score += 0.05
        elif bold_count > 0:
            print(f"PARTIAL: Only {bold_count}/4 row 3 cells are bold")
            total_score += 0.025
        else:
            print(f"FAIL: Row 3 header cells (A3:D3) are not bold")
    except Exception as e:
        print(f"ERROR: Component 6 (row 3 bold) — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score:.4f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
