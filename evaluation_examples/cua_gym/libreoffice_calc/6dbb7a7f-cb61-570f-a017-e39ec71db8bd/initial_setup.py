"""
Initial Setup: Equipment Loan Comparison
Task ID: calc_fin_loan_comparison_053
Domain: libreoffice_calc

Creates a spreadsheet with loan comparison data (pre-task state).
Rows 7-10 (Monthly Payment, Total Payments, Total Interest, Total Cost) are intentionally empty.
No formulas, no currency formatting, no conditional formatting, no sheet protection.
"""

import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_loan_comparison_053'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: LoanComparison ---
    ws = wb.active
    ws.title = 'LoanComparison'

    # Row 1: Title
    ws['A1'] = 'Equipment Loan Comparison'

    # Row 3: Headers (NOT bold — task requires making row 3 bold)
    headers = ['Parameter', 'Loan A', 'Loan C', 'Loan D']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)

    # Row 4: Loan Amount
    ws['A4'] = 'Loan Amount'
    ws['B4'] = 150000
    ws['C4'] = 150000
    ws['D4'] = 150000

    # Row 5: Annual Rate
    ws['A5'] = 'Annual Rate'
    ws['B5'] = 0.055
    ws['C5'] = 0.048
    ws['D5'] = 0.062

    # Row 6: Term (Years)
    ws['A6'] = 'Term (Years)'
    ws['B6'] = 5
    ws['C6'] = 7
    ws['D6'] = 4

    # Row 7: Monthly Payment — intentionally EMPTY (task will add PMT formulas)
    ws['A7'] = 'Monthly Payment'
    # B7, C7, D7 remain empty

    # Row 8: Total Payments — intentionally EMPTY (task will add formulas)
    ws['A8'] = 'Total Payments'
    # B8, C8, D8 remain empty

    # Row 9: Total Interest — intentionally EMPTY (task will add formulas)
    ws['A9'] = 'Total Interest'
    # B9, C9, D9 remain empty

    # Row 10: Total Cost — intentionally EMPTY (task will add formulas)
    ws['A10'] = 'Total Cost'
    # B10, C10, D10 remain empty

    # Column width adjustments for readability
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: LoanComparison')
    print('Rows 7-10 payment/summary cells: EMPTY (no formulas)')
    print('No formatting, no conditional formatting, no sheet protection')


create_initial()
