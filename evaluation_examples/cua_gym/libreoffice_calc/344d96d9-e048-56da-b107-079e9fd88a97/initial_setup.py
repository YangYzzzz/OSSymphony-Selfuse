"""
Initial Setup: Add comments to document locked vs unlocked cells before sheet protection
Task ID: calc_gsi_091
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_091'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Field Report Template ---
    ws = wb.active
    ws.title = "Field Report Template"

    # Styling
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    label_font = Font(name="Calibri", size=11, bold=True)
    input_fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 28

    # Row 1: Title (merged, locked)
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "Field Agent Weekly Report"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="FF1F3864", end_color="FF1F3864", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.protection = Protection(locked=True)

    # Row 2: Subtitle (merged, locked)
    ws.merge_cells("A2:D2")
    sub_cell = ws["A2"]
    sub_cell.value = "Complete all highlighted fields before submission"
    sub_cell.font = Font(name="Calibri", size=10, italic=True, color="666666")
    sub_cell.alignment = Alignment(horizontal="center")
    sub_cell.protection = Protection(locked=True)

    # Row 3: blank spacer
    ws.row_dimensions[3].height = 8

    # Row 4: Section header - Agent Information
    ws.merge_cells("A4:D4")
    sec1 = ws["A4"]
    sec1.value = "Agent Information"
    sec1.font = header_font
    sec1.fill = header_fill
    sec1.alignment = Alignment(horizontal="left")
    sec1.protection = Protection(locked=True)

    # Rows 5-8: Labels (locked) + Input cells (unlocked)
    labels_section1 = [
        ("A5", "Agent Name:", "B5"),
        ("C5", "Agent ID:", "D5"),
        ("A6", "Region:", "B6"),
        ("C6", "Report Date:", "D6"),
        ("A7", "Supervisor:", "B7"),
        ("C7", "Department:", "D7"),
        ("A8", "Contact Phone:", "B8"),
        ("C8", "Email:", "D8"),
    ]

    for label_coord, label_text, input_coord in labels_section1:
        lc = ws[label_coord]
        lc.value = label_text
        lc.font = label_font
        lc.border = thin_border
        lc.alignment = Alignment(vertical="center")
        lc.protection = Protection(locked=True)

        ic = ws[input_coord]
        ic.fill = input_fill
        ic.border = thin_border
        ic.protection = Protection(locked=False)

    # Row 9: blank spacer
    ws.row_dimensions[9].height = 8

    # Row 10: Section header - Weekly Metrics
    ws.merge_cells("A10:D10")
    sec2 = ws["A10"]
    sec2.value = "Weekly Performance Metrics"
    sec2.font = header_font
    sec2.fill = header_fill
    sec2.alignment = Alignment(horizontal="left")
    sec2.protection = Protection(locked=True)

    # Row 11: Sub-headers (locked)
    metric_headers = [
        ("A11", "Metric"),
        ("B11", "This Week"),
        ("C11", "Last Week"),
        ("D11", "Variance"),
    ]
    for coord, text in metric_headers:
        c = ws[coord]
        c.value = text
        c.font = Font(name="Calibri", size=11, bold=True, color="2F5496")
        c.fill = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")
        c.protection = Protection(locked=True)

    # Rows 12-16: Metric labels (locked), This Week & Last Week (unlocked input), Variance (locked formula)
    metrics = [
        ("Clients Visited", 12),
        ("New Leads Generated", 13),
        ("Contracts Signed", 14),
        ("Revenue Collected ($)", 15),
        ("Kilometers Traveled", 16),
    ]

    for metric_name, row in metrics:
        # A: metric label (locked)
        a = ws.cell(row=row, column=1, value=metric_name)
        a.font = Font(name="Calibri", size=11)
        a.border = thin_border
        a.protection = Protection(locked=True)

        # B: this week input (unlocked)
        b = ws.cell(row=row, column=2)
        b.fill = input_fill
        b.border = thin_border
        b.number_format = '#,##0' if 'Revenue' not in metric_name else '$#,##0.00'
        b.protection = Protection(locked=False)

        # C: last week input (unlocked)
        c = ws.cell(row=row, column=3)
        c.fill = input_fill
        c.border = thin_border
        c.number_format = '#,##0' if 'Revenue' not in metric_name else '$#,##0.00'
        c.protection = Protection(locked=False)

        # D: variance formula (locked)
        d = ws.cell(row=row, column=4, value=f'=B{row}-C{row}')
        d.border = thin_border
        d.number_format = '#,##0' if 'Revenue' not in metric_name else '$#,##0.00'
        d.protection = Protection(locked=True)

    # Row 17: blank spacer
    ws.row_dimensions[17].height = 8

    # Row 18: Section header - Notes
    ws.merge_cells("A18:D18")
    sec3 = ws["A18"]
    sec3.value = "Field Notes & Observations"
    sec3.font = header_font
    sec3.fill = header_fill
    sec3.alignment = Alignment(horizontal="left")
    sec3.protection = Protection(locked=True)

    # Rows 19-21: Notes area (unlocked)
    ws.merge_cells("A19:D21")
    notes = ws["A19"]
    notes.fill = input_fill
    notes.border = thin_border
    notes.alignment = Alignment(vertical="top", wrap_text=True)
    notes.protection = Protection(locked=False)

    # Row 22: blank spacer
    ws.row_dimensions[22].height = 8

    # Row 23: Footer formula row (locked)
    ws.merge_cells("A23:B23")
    footer_label = ws["A23"]
    footer_label.value = "Total Metrics Entered:"
    footer_label.font = Font(name="Calibri", size=11, bold=True)
    footer_label.protection = Protection(locked=True)

    ws.merge_cells("C23:D23")
    footer_val = ws["C23"]
    footer_val.value = '=COUNTA(B12:B16)+COUNTA(C12:C16)'
    footer_val.font = Font(name="Calibri", size=11, bold=True, color="2F5496")
    footer_val.protection = Protection(locked=True)

    # NO comments added - that is the agent's task
    # NO sheet protection applied - agent adds comments BEFORE protection

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
