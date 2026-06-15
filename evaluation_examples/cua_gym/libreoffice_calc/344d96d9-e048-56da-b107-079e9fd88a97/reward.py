"""
Reward Script: Add comments to document locked vs unlocked cells before sheet protection
Task ID: calc_gsi_091
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Locked cells have 'Protected - do not modify' comments
  Component 2 (0.30): Unlocked input cells have 'Enter value here' comments
  Component 3 (0.20): Comment author is consistent across all comments
  Component 4 (0.20): Sufficient comment coverage (minimum threshold of commented cells)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_091'


def persist_app_state(domain: str):
    """Try to save any unsaved changes via Ctrl+S."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify that cell comments have been added to document the protection scheme.
    Locked cells should have 'Protected - do not modify' and unlocked input cells
    should have 'Enter value here'.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get the first sheet (Field Report Template)
    try:
        ws = wb.worksheets[0]
        sheet_name = ws.title
        print(f"Working with sheet: {sheet_name}")
    except Exception as e:
        print(f"CRITICAL: Cannot access worksheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Categorize cells into locked and unlocked groups
    locked_cells_with_value = []  # cells that have content and are locked
    unlocked_input_cells = []     # cells that are unlocked (input fields)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            is_locked = True  # default in openpyxl
            if cell.protection:
                is_locked = cell.protection.locked

            if cell.value is not None and is_locked:
                locked_cells_with_value.append(cell)
            elif not is_locked and cell.value is None:
                # Unlocked empty cells are input fields
                unlocked_input_cells.append(cell)
            elif not is_locked and cell.value is not None:
                # Edge case: unlocked cells with values could be input cells too
                unlocked_input_cells.append(cell)

    print(f"Found {len(locked_cells_with_value)} locked cells with values")
    print(f"Found {len(unlocked_input_cells)} unlocked cells")

    # Collect ALL comments for later checks
    all_comments = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.comment:
                all_comments.append(cell)

    total_comment_count = len(all_comments)
    print(f"Total comments in sheet: {total_comment_count}")

    # If there are no comments at all, this is the initial state — score 0
    if total_comment_count == 0:
        print("FAIL: No comments found at all — this is the initial (unmodified) state")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Locked cells have 'Protected - do not modify' comments (0.30 points)
    # Check that locked cells with values have the protective comment
    try:
        locked_with_correct_comment = 0
        locked_checked = 0
        for cell in locked_cells_with_value:
            locked_checked += 1
            if cell.comment and 'protect' in cell.comment.text.lower() and 'not' in cell.comment.text.lower() and 'modif' in cell.comment.text.lower():
                locked_with_correct_comment += 1

        if locked_checked > 0:
            locked_ratio = locked_with_correct_comment / locked_checked
            if locked_ratio >= 0.8:
                print(f"PASS: Component 1 — {locked_with_correct_comment}/{locked_checked} locked cells have protective comment (0.30 pts)")
                total_score += 0.30
            elif locked_ratio >= 0.5:
                partial = round(0.30 * locked_ratio, 2)
                print(f"PARTIAL: Component 1 — {locked_with_correct_comment}/{locked_checked} locked cells have protective comment ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 1 — only {locked_with_correct_comment}/{locked_checked} locked cells have protective comment")
        else:
            print(f"FAIL: Component 1 — no locked cells with values found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Unlocked input cells have 'Enter value here' comments (0.30 points)
    # Check that unlocked cells have the input prompt comment
    try:
        # Focus on the key unlocked input cells that should have comments
        # (B5,D5, B6,D6, B7,D7, B8,D8, B12:C16, A19)
        key_input_coords = [
            'B5', 'D5', 'B6', 'D6', 'B7', 'D7', 'B8', 'D8',
            'B12', 'C12', 'B13', 'C13', 'B14', 'C14', 'B15', 'C15', 'B16', 'C16',
            'A19'
        ]
        unlocked_with_correct_comment = 0
        unlocked_checked = 0
        for coord in key_input_coords:
            cell = ws[coord]
            is_unlocked = cell.protection and not cell.protection.locked
            if is_unlocked:
                unlocked_checked += 1
                if cell.comment and 'enter' in cell.comment.text.lower() and 'value' in cell.comment.text.lower():
                    unlocked_with_correct_comment += 1

        if unlocked_checked > 0:
            unlocked_ratio = unlocked_with_correct_comment / unlocked_checked
            if unlocked_ratio >= 0.8:
                print(f"PASS: Component 2 — {unlocked_with_correct_comment}/{unlocked_checked} key unlocked cells have input prompt comment (0.30 pts)")
                total_score += 0.30
            elif unlocked_ratio >= 0.5:
                partial = round(0.30 * unlocked_ratio, 2)
                print(f"PARTIAL: Component 2 — {unlocked_with_correct_comment}/{unlocked_checked} key unlocked cells have input prompt comment ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 2 — only {unlocked_with_correct_comment}/{unlocked_checked} key unlocked cells have input prompt comment")
        else:
            print(f"FAIL: Component 2 — no key unlocked input cells found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Comment author consistency (0.20 points)
    # All comments should have a consistent author name (not empty)
    try:
        authors = set()
        for cell in all_comments:
            if cell.comment.author:
                authors.add(cell.comment.author.strip())

        if len(authors) == 1:
            author_name = list(authors)[0]
            if len(author_name) > 0:
                print(f"PASS: Component 3 — all comments have consistent author '{author_name}' (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 3 — author name is empty")
        elif len(authors) > 1:
            # Partial credit if most comments share same author
            from collections import Counter
            author_counts = Counter()
            for cell in all_comments:
                if cell.comment.author:
                    author_counts[cell.comment.author.strip()] += 1
            most_common = author_counts.most_common(1)[0]
            ratio = most_common[1] / total_comment_count
            if ratio >= 0.8:
                print(f"PARTIAL: Component 3 — {most_common[1]}/{total_comment_count} comments share author '{most_common[0]}' (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 3 — inconsistent authors: {authors}")
        else:
            print(f"FAIL: Component 3 — no authors found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Sufficient comment coverage (0.20 points)
    # The task requires comments on BOTH locked and unlocked cells
    # Golden has comments on ~48 cells. We check that there are enough comments
    # to indicate both locked and unlocked cells were documented.
    try:
        # Count comments on locked vs unlocked cells
        comments_on_locked = 0
        comments_on_unlocked = 0
        for cell in all_comments:
            is_locked = True
            if cell.protection:
                is_locked = cell.protection.locked
            if is_locked:
                comments_on_locked += 1
            else:
                comments_on_unlocked += 1

        has_both_types = comments_on_locked >= 5 and comments_on_unlocked >= 5
        has_good_coverage = total_comment_count >= 20

        if has_both_types and has_good_coverage:
            print(f"PASS: Component 4 — {comments_on_locked} comments on locked, {comments_on_unlocked} on unlocked, {total_comment_count} total (0.20 pts)")
            total_score += 0.20
        elif has_both_types:
            print(f"PARTIAL: Component 4 — both types commented but low coverage ({total_comment_count} total) (0.10 pts)")
            total_score += 0.10
        elif comments_on_locked >= 3 or comments_on_unlocked >= 3:
            print(f"PARTIAL: Component 4 — only partial type coverage: locked={comments_on_locked}, unlocked={comments_on_unlocked} (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4 — insufficient coverage: locked={comments_on_locked}, unlocked={comments_on_unlocked}, total={total_comment_count}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
