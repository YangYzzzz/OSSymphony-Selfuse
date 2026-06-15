"""
FINAL REWARD SCRIPT - SUCCESS
Task: I need a new column called "Conversion Rate (%)" that calculates the conversion rates from my data. Convert to number format and mark the highest rate in red (#ff0000) font.
Generated: 2025-11-24 07:26:01
Status: success
Model: o3
Total Steps: 1
"""

import openpyxl, re, os

def verify_conversion_rate_task(file_path: str) -> float:
    """Verify the LibreOffice/Excel task:
    1. A new column called "Conversion Rate (%)" exists.
    2. It correctly calculates Conversions / Visitors for each data row (either as numeric value or via formula).
    3. Cells are formatted as percentage (number format containing '%').
    4. The highest conversion-rate cell is displayed in red font (#ff0000).

    Returns a progressive score between 0.0 and 1.0.
    """

    print(f"Starting verification for file: {file_path}")

    max_score = 1.0
    score = 0.0

    # Scoring weights (must add up to 1.0)
    SCORE_COLUMN   = 0.1  # Column presence
    SCORE_CALC     = 0.5  # Correct calculations / formulas
    SCORE_NUMFMT   = 0.1  # Percentage formatting
    SCORE_COLOR    = 0.3  # Highest value red font

    # --- Load workbook twice: once with values, once with formulas ---
    try:
        wb_vals     = openpyxl.load_workbook(file_path, data_only=True)
        wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
        print("✓ Workbook loaded successfully (values & formulas)")
    except Exception as e:
        print(f"✗ Unable to load workbook: {e}")
        return 0.0

    sheet_name = 'Sheet' if 'Sheet' in wb_vals.sheetnames else wb_vals.sheetnames[0]
    sh_vals     = wb_vals[sheet_name]
    sh_formulas = wb_formulas[sheet_name]
    print(f"Using sheet: {sheet_name}")

    # ------------------------------------------------------------------
    # 1) Verify column presence
    # ------------------------------------------------------------------
    headers = [cell.value for cell in sh_vals[1]]
    print(f"Headers found: {headers}")

    # Locate essential columns
    try:
        idx_visitors   = headers.index('Visitors')
        idx_convs      = headers.index('Conversions')
    except ValueError:
        print("✗ Missing essential headers 'Visitors' or 'Conversions'")
        return 0.0

    if 'Conversion Rate (%)' in headers:
        idx_rate = headers.index('Conversion Rate (%)')
        print("✓ 'Conversion Rate (%)' column present")
        score += SCORE_COLUMN
    else:
        print("✗ 'Conversion Rate (%)' column NOT found")
        return round(score, 2)  # Cannot perform deeper checks without the column

    # ------------------------------------------------------------------
    # 2) Verify calculations & number formats
    # ------------------------------------------------------------------
    calc_correct   = True
    numfmt_correct = True
    rates          = []
    max_rate_val   = -1
    max_rate_row   = None

    for row_idx, row in enumerate(sh_vals.iter_rows(min_row=2, values_only=False), start=2):
        # Fetch numeric visitors / conversions
        try:
            visitors    = float(row[idx_visitors].value)
            conversions = float(row[idx_convs].value)
        except (TypeError, ValueError):
            print(f"✗ Non-numeric Visitors/Conversions in row {row_idx}")
            calc_correct = False
            break

        expected_rate = conversions / visitors if visitors else None
        if expected_rate is None:
            print(f"✗ Zero Visitors in row {row_idx}")
            calc_correct = False
            break

        # Observed value & formula
        rate_val_cell   = row[idx_rate]
        rate_formula_cell = sh_formulas.cell(row=row_idx, column=idx_rate + 1)

        observed_val = rate_val_cell.value
        numeric_match  = isinstance(observed_val, (int, float)) and abs(observed_val - expected_rate) < 1e-4

        # If value mismatch, inspect formula (should reference same-row Conversions/Visitors)
        formula_match = False
        if not numeric_match and rate_formula_cell.value and str(rate_formula_cell.value).startswith('='):
            conv_col_letter = openpyxl.utils.get_column_letter(idx_convs + 1)
            vis_col_letter  = openpyxl.utils.get_column_letter(idx_visitors + 1)
            pat1 = rf"{conv_col_letter}{row_idx}.*\/{vis_col_letter}{row_idx}"
            pat2 = rf"{vis_col_letter}{row_idx}.*\/.*{conv_col_letter}{row_idx}"
            formula_match = bool(re.search(pat1, rate_formula_cell.value.replace('$', ''), re.IGNORECASE) or
                                  re.search(pat2, rate_formula_cell.value.replace('$', ''), re.IGNORECASE))

        if not (numeric_match or formula_match):
            print(f"✗ Incorrect conversion-rate calculation in row {row_idx}")
            calc_correct = False
            break

        # Number-format check (percent expected)
        numfmt = rate_formula_cell.number_format or rate_val_cell.number_format
        if '%' not in str(numfmt):
            numfmt_correct = False

        # Track highest rate value
        if expected_rate > max_rate_val:
            max_rate_val = expected_rate
            max_rate_row = row_idx

    if calc_correct:
        print("✓ Conversion rate calculations appear correct for all rows")
        score += SCORE_CALC
    else:
        print("✗ Conversion rate calculation errors detected")

    if numfmt_correct:
        print("✓ Cells formatted as percentage")
        score += SCORE_NUMFMT
    else:
        print("✗ Some cells lack proper percentage formatting")

    # ------------------------------------------------------------------
    # 3) Verify highest rate highlighted in red font
    # ------------------------------------------------------------------
    highest_color_correct = False
    if max_rate_row is not None:
        font_color = sh_formulas.cell(row=max_rate_row, column=idx_rate + 1).font.color
        if font_color is not None and font_color.type == 'rgb' and font_color.rgb:
            if font_color.rgb.upper().endswith('FF0000'):
                highest_color_correct = True

    if highest_color_correct:
        print(f"✓ Highest rate (row {max_rate_row}) is red (#ff0000)")
        score += SCORE_COLOR
    else:
        print(f"✗ Highest rate (row {max_rate_row}) is not red as required")

    final_score = round(min(score, max_score), 2)
    print(f"Total Score: {final_score} / {max_score}")
    return final_score


if __name__ == "__main__":
    target_file = "/home/user/i_need_a_new_column_called_conversion_rate_that_calculates_the_conversion_rates_from_my_data_convert.xlsx"
    reward = verify_conversion_rate_task(target_file)
    print(f"REWARD: {reward}")
