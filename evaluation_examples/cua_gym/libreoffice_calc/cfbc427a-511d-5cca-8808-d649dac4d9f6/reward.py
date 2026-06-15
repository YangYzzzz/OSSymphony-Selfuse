"""
Reward Script: Paste Special with Link — live link from Dashboard to Data sheet
Task ID: calc_gsi_053
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): Dashboard contains formulas referencing Data sheet
  Component 2 (0.25): Header row linked (all 7 columns reference Data row 1)
  Component 3 (0.30): Data rows linked (at least 14 of 15 employee rows linked across columns)
  Component 4 (0.10): Full coverage — all 7 columns x 16 rows linked (112 cells)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_053'


def verify_task(file_path):
    """
    Verify that the Dashboard sheet contains live links (=Data!XX formulas)
    to the Data sheet, created via Paste Special > Link.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: both sheets must exist
    if 'Data' not in wb.sheetnames:
        print("FAIL: 'Data' sheet not found")
        print("REWARD: 0.0")
        return 0.0
    if 'Dashboard' not in wb.sheetnames:
        print("FAIL: 'Dashboard' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws_dashboard = wb['Dashboard']

    # Helper: check if a cell value is a formula referencing the Data sheet
    def is_data_ref(cell_value):
        """Returns True if the cell contains a formula like =Data!XX"""
        if not isinstance(cell_value, str):
            return False
        # Match patterns like =Data!A1, =Data!$A$1, ='Data'!A1
        return bool(re.match(r"^=('?Data'?!)", cell_value))

    # Scan Dashboard for all Data-referencing formulas
    # We look in rows 2..30 (generous range) to find where the links are
    data_ref_cells = []
    for row in ws_dashboard.iter_rows(min_row=2, max_row=30, min_col=1, max_col=10):
        for cell in row:
            if cell.value is not None and is_data_ref(str(cell.value)):
                data_ref_cells.append(cell.coordinate)

    print(f"INFO: Found {len(data_ref_cells)} cells with =Data! references in Dashboard")

    # Component 1: Dashboard contains formulas referencing Data sheet (0.35 points)
    # The initial file has ZERO such references; golden has 112 (7 cols x 16 rows)
    try:
        if len(data_ref_cells) >= 7:
            print(f"PASS: Component 1 — Dashboard has {len(data_ref_cells)} Data references (>= 7) (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 1 — Dashboard has only {len(data_ref_cells)} Data references (need >= 7)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Header row linked (0.25 points)
    # The golden file links headers in row 3: =Data!A1, =Data!B1, ..., =Data!G1
    # We check that at least one row has all 7 columns referencing Data row 1
    try:
        header_row_found = 0
        for check_row in range(2, 25):
            header_refs = 0
            for col_idx in range(1, 8):  # columns A-G
                cell_val = ws_dashboard.cell(row=check_row, column=col_idx).value
                if cell_val and is_data_ref(str(cell_val)):
                    # Check it references row 1 of Data sheet
                    val_str = str(cell_val).upper()
                    if re.search(r"DATA'?![A-Z$]*1\b", val_str):
                        header_refs += 1
            if header_refs >= 6:  # at least 6 of 7 headers linked
                header_row_found = check_row
                break
        if header_row_found > 0:
            print(f"PASS: Component 2 — Header row linked at Dashboard row {header_row_found} (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — No row in Dashboard has header references to Data row 1")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Data rows linked (0.30 points)
    # Golden file has 15 employee rows (Data rows 2-16) linked in Dashboard.
    # We check how many unique Data source rows (2-16) are referenced.
    try:
        referenced_source_rows = set()
        for row in ws_dashboard.iter_rows(min_row=2, max_row=30, min_col=1, max_col=10):
            for cell in row:
                val = cell.value
                if val and is_data_ref(str(val)):
                    # Extract the row number from the reference, e.g., =Data!A2 -> 2
                    match = re.search(r"DATA'?!\$?[A-Z]+\$?(\d+)", str(val).upper())
                    if match:
                        src_row = int(match.group(1))
                        if 2 <= src_row <= 16:
                            referenced_source_rows.add(src_row)

        num_linked_rows = len(referenced_source_rows)
        print(f"INFO: {num_linked_rows} of 15 Data source rows (2-16) are referenced in Dashboard")

        if num_linked_rows >= 14:
            print(f"PASS: Component 3 — {num_linked_rows}/15 data rows linked (0.30 pts)")
            total_score += 0.30
        elif num_linked_rows >= 8:
            partial = round(0.30 * (num_linked_rows / 15), 2)
            print(f"PARTIAL: Component 3 — {num_linked_rows}/15 data rows linked ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {num_linked_rows}/15 data rows linked (need >= 14)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Full coverage — all 7 columns across all 16 source rows (0.10 points)
    # Golden has 112 cells (7*16). We require at least 100 to account for minor variations.
    try:
        if len(data_ref_cells) >= 100:
            print(f"PASS: Component 4 — Full coverage: {len(data_ref_cells)} linked cells (>= 100) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 — Only {len(data_ref_cells)} linked cells (need >= 100 for full coverage)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
