"""
Initial Setup: Paste Special with Link - live link from Dashboard to Data sheet
Task ID: calc_gsi_053
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_053'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Data (source sheet) ---
    ws_data = wb.active
    ws_data.title = 'Data'

    # Headers
    headers = ['Employee', 'Department', 'Q1 Sales', 'Q2 Sales', 'Q3 Sales', 'Q4 Sales', 'Annual Total']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Realistic employee sales data
    data = [
        ['Sarah Chen', 'Engineering', 45230, 52140, 48900, 61200, None],
        ['Marcus Johnson', 'Marketing', 38750, 41200, 39800, 44500, None],
        ['Priya Patel', 'Sales', 67800, 72350, 69100, 78400, None],
        ['James Rodriguez', 'Engineering', 41500, 43800, 45200, 47600, None],
        ['Aisha Williams', 'Finance', 35200, 36800, 38100, 39500, None],
        ['David Kim', 'Sales', 58900, 63200, 61400, 68700, None],
        ['Emily Thompson', 'Marketing', 42100, 44500, 43200, 46800, None],
        ['Robert Nakamura', 'Engineering', 49300, 51700, 53200, 55800, None],
        ['Lisa Martinez', 'Finance', 33800, 35100, 36400, 37200, None],
        ['Michael O\'Brien', 'Sales', 71200, 74500, 72800, 79300, None],
        ['Jennifer Chang', 'Marketing', 39600, 42100, 40800, 45200, None],
        ['Daniel Foster', 'Engineering', 46700, 48300, 50100, 52400, None],
        ['Rachel Green', 'Finance', 37100, 38500, 39800, 41200, None],
        ['Thomas Wright', 'Sales', 62400, 65800, 63700, 70100, None],
        ['Sophia Lee', 'Marketing', 41800, 43600, 42500, 47100, None],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_data.cell(row=r, column=c, value=val)
            if c >= 3 and c <= 6 and val is not None:
                cell.number_format = '$#,##0'

    # Add Annual Total formulas in column G
    for r in range(2, 17):
        cell = ws_data.cell(row=r, column=7, value=f'=SUM(C{r}:F{r})')
        cell.number_format = '$#,##0'

    # Set column widths
    ws_data.column_dimensions['A'].width = 22
    ws_data.column_dimensions['B'].width = 15
    for col_letter in ['C', 'D', 'E', 'F', 'G']:
        ws_data.column_dimensions[col_letter].width = 14

    # Freeze header row
    ws_data.freeze_panes = 'A2'

    # --- Sheet 2: Dashboard (destination - empty, waiting for linked data) ---
    ws_dash = wb.create_sheet('Dashboard')

    # Add a title to Dashboard
    title_cell = ws_dash.cell(row=1, column=1, value='Sales Dashboard')
    title_cell.font = Font(name='Calibri', size=14, bold=True, color='2F5496')
    ws_dash.merge_cells('A1:G1')
    title_cell.alignment = Alignment(horizontal='center')

    # Add instruction note in row 2
    note_cell = ws_dash.cell(row=2, column=1, value='(Link data from Data sheet below)')
    note_cell.font = Font(name='Calibri', size=10, italic=True, color='808080')
    ws_dash.merge_cells('A2:G2')
    note_cell.alignment = Alignment(horizontal='center')

    # Set column widths to match Data sheet
    ws_dash.column_dimensions['A'].width = 22
    ws_dash.column_dimensions['B'].width = 15
    for col_letter in ['C', 'D', 'E', 'F', 'G']:
        ws_dash.column_dimensions[col_letter].width = 14

    # Dashboard is intentionally empty below row 2 - no links, no references
    # The task is to use Paste Special > Link to create live references here

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
