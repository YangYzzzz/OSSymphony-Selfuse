"""
Initial Setup: Multi-app terminal + calc task — latitude.xlsx and longitude.ods on Desktop
Task ID: osworld_multi_apps_terminal_calc_013
Domain: libreoffice_calc (multi-app: terminal + calc)

Creates:
  - /home/user/Desktop/latitude.xlsx  (one column 'Lat' with 12 latitude values)
  - /home/user/Desktop/longitude.ods  (one column 'Lon' with 12 longitude values)
  - Opens a terminal (gnome-terminal) in the Desktop directory

Initial state does NOT contain coordinates.csv (agent must create it).
"""

import os
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user/Desktop'
TASK_ID = 'osworld_multi_apps_terminal_calc_013'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_latitude_xlsx():
    """Create latitude.xlsx on the Desktop with a 'Lat' header and 12 latitude values."""
    output = f'{WORKDIR}/latitude.xlsx'

    # Realistic latitudes (one per city, matching the longitude.ods order)
    latitudes = [
        1.3521,    # Singapore
        -33.8688,  # Sydney
        35.6762,   # Tokyo
        48.8566,   # Paris
        51.5074,   # London
        -22.9068,  # Rio de Janeiro
        40.7128,   # New York
        55.7558,   # Moscow
        -34.6037,  # Buenos Aires
        28.6139,   # New Delhi
        -1.2921,   # Nairobi
        19.4326,   # Mexico City
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Latitude'

    # Header
    ws.cell(row=1, column=1, value='Lat')

    # Data
    for i, lat in enumerate(latitudes, start=2):
        ws.cell(row=i, column=1, value=lat)

    wb.save(output)
    print(f'Created: {output}')


def create_longitude_ods():
    """Create longitude.ods on the Desktop with a 'Lon' header and 12 longitude values."""
    output = f'{WORKDIR}/longitude.ods'

    # Matching longitudes (same city order as latitudes)
    longitudes = [
        103.8198,  # Singapore
        151.2093,  # Sydney
        139.6503,  # Tokyo
        2.3522,    # Paris
        -0.1278,   # London
        -43.1729,  # Rio de Janeiro
        -74.0060,  # New York
        37.6173,   # Moscow
        -58.3816,  # Buenos Aires
        77.2090,   # New Delhi
        36.8219,   # Nairobi
        -99.1332,  # Mexico City
    ]

    # Build ODS content using a minimal XML structure (ODS is a zip of XML files)
    # We use Python to write a simple ODS file manually via the odfpy or pyexcel approach
    # Since odfpy may not be available, we use a subprocess call to LibreOffice to convert
    # from a CSV, or we write the ODS XML manually.
    # Simplest reliable approach: write a temp CSV and convert to ODS using LibreOffice.

    import tempfile
    import zipfile
    import io

    # Write minimal ODS (Open Document Spreadsheet) format
    # ODS is a ZIP file containing content.xml and a few other files
    content_xml = '''<?xml version="1.0" encoding="UTF-8"?>
<office:document-content
    xmlns:office="urn:oasis:names:tc:opendocument:xmlns:office:1.0"
    xmlns:table="urn:oasis:names:tc:opendocument:xmlns:table:1.0"
    xmlns:text="urn:oasis:names:tc:opendocument:xmlns:text:1.0"
    xmlns:fo="urn:oasis:names:tc:opendocument:xmlns:xsl-fo-compatible:1.0"
    xmlns:number="urn:oasis:names:tc:opendocument:xmlns:datastyle:1.0"
    xmlns:of="urn:oasis:names:tc:opendocument:xmlns:of:1.2"
    office:version="1.2">
  <office:scripts/>
  <office:automatic-styles>
    <number:number-style style:name="N0" xmlns:style="urn:oasis:names:tc:opendocument:xmlns:style:1.0">
      <number:number number:decimal-places="4" number:min-integer-digits="1"/>
    </number:number-style>
  </office:automatic-styles>
  <office:body>
    <office:spreadsheet>
      <table:table table:name="Longitude">
        <table:table-column table:number-columns-repeated="1"/>
        <table:table-row>
          <table:table-cell office:value-type="string">
            <text:p>Lon</text:p>
          </table:table-cell>
        </table:table-row>
'''
    for lon in longitudes:
        content_xml += f'''        <table:table-row>
          <table:table-cell office:value-type="float" office:value="{lon}">
            <text:p>{lon}</text:p>
          </table:table-cell>
        </table:table-row>
'''
    content_xml += '''      </table:table>
    </office:spreadsheet>
  </office:body>
</office:document-content>'''

    manifest_xml = '''<?xml version="1.0" encoding="UTF-8"?>
<manifest:manifest xmlns:manifest="urn:oasis:names:tc:opendocument:xmlns:manifest:1.0" manifest:version="1.2">
  <manifest:file-entry manifest:full-path="/" manifest:version="1.2" manifest:media-type="application/vnd.oasis.opendocument.spreadsheet"/>
  <manifest:file-entry manifest:full-path="content.xml" manifest:media-type="text/xml"/>
  <manifest:file-entry manifest:full-path="mimetype" manifest:media-type=""/>
</manifest:manifest>'''

    mimetype = 'application/vnd.oasis.opendocument.spreadsheet'

    meta_xml = '''<?xml version="1.0" encoding="UTF-8"?>
<office:document-meta xmlns:office="urn:oasis:names:tc:opendocument:xmlns:office:1.0" office:version="1.2">
  <office:meta/>
</office:document-meta>'''

    with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zf:
        # mimetype MUST be first and uncompressed
        zf.writestr(zipfile.ZipInfo('mimetype'), mimetype)
        zf.writestr('META-INF/manifest.xml', manifest_xml)
        zf.writestr('content.xml', content_xml)
        zf.writestr('meta.xml', meta_xml)

    print(f'Created: {output}')


def create_initial():
    # Ensure Desktop directory exists
    os.makedirs(WORKDIR, exist_ok=True)

    # Remove any existing coordinates.csv to ensure clean initial state
    csv_path = f'{WORKDIR}/coordinates.csv'
    if os.path.exists(csv_path):
        os.remove(csv_path)
        print(f'Removed existing: {csv_path}')

    # Create the two source files
    create_latitude_xlsx()
    create_longitude_ods()

    # GUI-ready startup: open a terminal in the Desktop directory
    # Use gnome-terminal with working directory set to Desktop
    launch_gui(
        'gnome-terminal --working-directory=/home/user/Desktop',
        delay_sec=2.0
    )
    print('GUI_READY: launched gnome-terminal on Desktop with DISPLAY=:0')


create_initial()
