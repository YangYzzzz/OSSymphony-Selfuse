"""
Initial Setup: Horizontal lookup table with shifted row index issue
Task ID: calc_tbl_058
Domain: libreoffice_calc

Creates a pricing lookup spreadsheet where HLOOKUP formulas reference
row_index 3 but data has shifted down due to an inserted row, so they
now return product names instead of prices.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_058'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pricing"

    # Styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    data_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # === HORIZONTAL TABLE (rows 1-6) ===
    # Row 1: Category headers
    categories = [
        "Category", "Electronics", "Furniture", "Clothing",
        "Groceries", "Sports", "Books"
    ]
    for c, cat in enumerate(categories, 1):
        cell = ws.cell(row=1, column=c, value=cat)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Row 2: INSERTED row (this is the new row that shifted everything down)
    ws.cell(row=2, column=1, value="Region").font = Font(bold=True, italic=True)
    regions = ["North America", "Europe", "Asia Pacific", "Latin America", "Middle East", "Global"]
    for c, region in enumerate(regions, 2):
        cell = ws.cell(row=2, column=c, value=region)
        cell.font = Font(italic=True, color="808080")
        cell.alignment = data_align
        cell.border = thin_border

    # Row 3: Product Names (was originally row 2 before insertion)
    ws.cell(row=3, column=1, value="Product").font = Font(bold=True)
    products = [
        "Samsung Galaxy S24", "Herman Miller Aeron", "Nike Air Max 90",
        "Organic Quinoa 5lb", "Titleist Pro V1 Golf Balls", "Sapiens by Harari"
    ]
    for c, prod in enumerate(products, 2):
        cell = ws.cell(row=3, column=c, value=prod)
        cell.alignment = data_align
        cell.border = thin_border

    # Row 4: Prices (was originally row 3 before insertion - THIS IS THE TARGET DATA)
    ws.cell(row=4, column=1, value="Price ($)").font = Font(bold=True)
    prices = [849.99, 1395.00, 129.95, 24.50, 54.99, 18.95]
    for c, price in enumerate(prices, 2):
        cell = ws.cell(row=4, column=c, value=price)
        cell.number_format = '$#,##0.00'
        cell.alignment = data_align
        cell.border = thin_border

    # Row 5: Stock quantity (was originally row 4)
    ws.cell(row=5, column=1, value="Stock").font = Font(bold=True)
    stock = [2450, 180, 3200, 8500, 1100, 15000]
    for c, qty in enumerate(stock, 2):
        cell = ws.cell(row=5, column=c, value=qty)
        cell.number_format = '#,##0'
        cell.alignment = data_align
        cell.border = thin_border

    # Row 6: Supplier (was originally row 5)
    ws.cell(row=6, column=1, value="Supplier").font = Font(bold=True)
    suppliers = [
        "TechDist Corp", "Office Furnish Ltd", "Athletic Wholesale",
        "FreshOrg Supply", "ProGolf Distrib", "Penguin Random House"
    ]
    for c, sup in enumerate(suppliers, 2):
        cell = ws.cell(row=6, column=c, value=sup)
        cell.alignment = data_align
        cell.border = thin_border

    # Row 8: Lookup section header
    ws.cell(row=8, column=1, value="Price Lookup").font = Font(size=13, bold=True, color="2F5496")
    ws.merge_cells("A8:H8")
    ws["A8"].alignment = Alignment(horizontal="center")

    # Row 9: Lookup labels
    ws.cell(row=9, column=1, value="Lookup Category").font = Font(bold=True)
    ws.cell(row=9, column=2, value="Lookup Key").font = Font(bold=True)
    lookup_cats = ["Electronics", "Furniture", "Clothing", "Groceries", "Sports", "Books"]
    for c, cat in enumerate(lookup_cats, 3):
        cell = ws.cell(row=9, column=c, value=cat)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
        cell.alignment = header_align
        cell.border = thin_border

    # Row 10: HLOOKUP formulas - THESE USE row_index_num=3 (WRONG!)
    # After the row insertion, the prices are now in row 4 of the table,
    # so row_index_num should be 4, but these still say 3
    ws.cell(row=10, column=1, value="Result").font = Font(bold=True)
    ws.cell(row=10, column=2, value="Price").font = Font(bold=True)

    # HLOOKUP(lookup_value, table_array, row_index_num, [range_lookup])
    # table_array is A1:G6 (the full horizontal table)
    # row_index_num=3 points to row 3 of the table (product names) - WRONG
    # Should be row_index_num=4 to get prices
    for c in range(3, 9):  # C10 through H10
        # lookup_value is the category in row 9
        col_letter = openpyxl.utils.get_column_letter(c)
        formula = f'=HLOOKUP({col_letter}9,$A$1:$G$6,3,FALSE)'
        cell = ws.cell(row=10, column=c, value=formula)
        cell.number_format = '$#,##0.00'
        cell.alignment = data_align
        cell.border = thin_border
        # Yellow background to highlight the lookup results
        cell.fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")

    # Row 12: Note explaining the issue
    ws.cell(row=12, column=1, value="Note:").font = Font(bold=True, color="FF0000")
    ws.cell(row=12, column=2, value="A new 'Region' row was inserted at row 2, shifting data down. "
            "The HLOOKUP formulas in row 10 still use the old row index.").font = Font(color="FF0000")

    # Column widths
    ws.column_dimensions["A"].width = 18
    for col_letter in ["B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col_letter].width = 22

    # Freeze panes
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
