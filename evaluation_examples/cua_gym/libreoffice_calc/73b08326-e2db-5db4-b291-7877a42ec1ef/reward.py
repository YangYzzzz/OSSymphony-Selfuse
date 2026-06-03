"""
Reward Script: Update HLOOKUP row index from 3 to 4 in row 10 formulas
Task ID: calc_tbl_058
Domain: libreoffice_calc
Scoring:
  Component 1 (0.7 pts): HLOOKUP formulas in C10:G10 all use row_index 4 with correct structure
  Component 2 (0.3 pts): HLOOKUP formula in H10 also uses row_index 4 with correct structure
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_058'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def check_hlookup_updated(formula):
    """
    Check if a formula is a valid HLOOKUP with row_index 4.
    Returns True only if:
      - It's an HLOOKUP formula
      - The row_index (3rd arg) is 4
      - It uses FALSE for exact match
    """
    if not formula or not isinstance(formula, str):
        return False
    formula_clean = formula.replace(' ', '')
    formula_upper = formula_clean.upper()
    # Must be HLOOKUP with row_index 4 and FALSE
    if 'HLOOKUP(' not in formula_upper:
        return False
    if 'FALSE' not in formula_upper:
        return False
    # Extract row_index (3rd argument)
    m = re.match(r'=HLOOKUP\(.+?,\s*.+?,\s*(\d+)\s*,', formula_clean, re.IGNORECASE)
    if m and int(m.group(1)) == 4:
        return True
    return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Component 1: HLOOKUP formulas in C10:G10 updated to row_index 4 (0.7 points)
    # Initial has row_index 3 (broken), golden has 4 (fixed).
    # Each of the 5 cells contributes 0.14 points for partial credit.
    try:
        cols_main = ['C', 'D', 'E', 'F', 'G']
        pass_count = 0
        for col in cols_main:
            cell_ref = f'{col}10'
            formula = ws[cell_ref].value
            if check_hlookup_updated(formula):
                pass_count += 1
                print(f"  PASS: {cell_ref} — HLOOKUP with row_index 4 (formula: {formula})")
            else:
                print(f"  FAIL: {cell_ref} — expected HLOOKUP row_index 4, found: {formula}")

        if pass_count == len(cols_main):
            print(f"PASS: Component 1 — All {len(cols_main)} HLOOKUP formulas in C10:G10 updated (0.7 pts)")
            total_score += 0.7
        elif pass_count > 0:
            partial = round(0.7 * pass_count / len(cols_main), 2)
            print(f"PARTIAL: Component 1 — {pass_count}/{len(cols_main)} updated ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No HLOOKUP formulas updated to row_index 4")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: H10 HLOOKUP formula also updated to row_index 4 (0.3 points)
    # H10 is the 6th HLOOKUP in the range, looking up "Books"
    try:
        formula_h10 = ws['H10'].value
        if check_hlookup_updated(formula_h10):
            print(f"PASS: Component 2 — H10 HLOOKUP updated to row_index 4 (formula: {formula_h10}) (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — H10 expected HLOOKUP row_index 4, found: {formula_h10}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
