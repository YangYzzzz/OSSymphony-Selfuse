"""
Initial Setup: Add conditional formatting to highlight weekend dates in orange
Task ID: osworld_calc_conditional_format_weekday_002
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from datetime import date

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_conditional_format_weekday_002'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Schedule ---
    ws = wb.active
    ws.title = "Schedule"

    # Headers
    headers = ["Task Name", "Date", "Assignee", "Status"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic project schedule data
    # Dates are a mix of weekdays and weekends (2025-03)
    # Note: 2025-03-01 is Saturday, 2025-03-02 is Sunday
    data = [
        ["Project Kickoff",          date(2025, 3, 3),  "Sarah Chen",      "Completed"],
        ["Requirements Review",      date(2025, 3, 4),  "Marcus Johnson",  "Completed"],
        ["Architecture Design",      date(2025, 3, 5),  "Elena Rodriguez", "In Progress"],
        ["Weekend Sprint Planning",  date(2025, 3, 8),  "David Kim",       "Pending"],
        ["Database Schema Draft",    date(2025, 3, 9),  "Sarah Chen",      "Pending"],
        ["API Design Review",        date(2025, 3, 11), "Priya Patel",     "Pending"],
        ["Weekend Code Review",      date(2025, 3, 15), "Marcus Johnson",  "Pending"],
        ["Frontend Mockups",         date(2025, 3, 16), "Elena Rodriguez", "Pending"],
        ["Backend Implementation",   date(2025, 3, 17), "David Kim",       "Pending"],
        ["QA Testing Start",         date(2025, 3, 22), "Priya Patel",     "Pending"],
        ["Weekend Integration Test", date(2025, 3, 23), "Sarah Chen",      "Pending"],
        ["Bug Fixing Sprint",        date(2025, 3, 24), "Marcus Johnson",  "Pending"],
        ["Performance Optimization", date(2025, 3, 26), "Elena Rodriguez", "Pending"],
        ["User Acceptance Testing",  date(2025, 3, 28), "David Kim",       "Pending"],
        ["Product Launch",           date(2025, 3, 31), "Priya Patel",     "Pending"],
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        ws.cell(row=r, column=2, value=row_data[1])
        ws.cell(row=r, column=3, value=row_data[2])
        ws.cell(row=r, column=4, value=row_data[3])

    # Set date format for column B
    for r in range(2, 17):
        ws.cell(row=r, column=2).number_format = 'yyyy-mm-dd'

    # Set column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
