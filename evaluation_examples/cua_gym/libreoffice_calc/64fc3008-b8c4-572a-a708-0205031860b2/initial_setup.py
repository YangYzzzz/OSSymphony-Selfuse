"""
Initial Setup: arxiv_papers.xlsx with two ArXiv paper entries
Task ID: osworld_multi_apps_pdf_download_cite_001
Domain: libreoffice_calc (multi-app: also opens Chrome)

Creates:
  - /home/user/arxiv_papers.xlsx: spreadsheet listing two ArXiv papers with links
  - Launches LibreOffice Calc with the spreadsheet open
  - Launches Chrome for downloading the PDF

MUST NOT include: paper01.pdf or citation_check.docx (those are produced by the agent)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_pdf_download_cite_001'
OUTPUT = f'{WORKDIR}/arxiv_papers.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ArXiv Papers"

    # --- Header row styling ---
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Column headers
    headers = ["Title", "ArXiv Link"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # --- Paper data (two ArXiv papers) ---
    data = [
        [
            "Attention Is All You Need",
            "https://arxiv.org/abs/1706.03762"
        ],
        [
            "BERT: Pre-training of Deep Bidirectional Transformers for Language Understanding",
            "https://arxiv.org/abs/1810.04805"
        ],
    ]

    data_font = Font(name="Calibri", size=11)
    data_align_title = Alignment(horizontal="left", vertical="center", wrap_text=True)
    data_align_link = Alignment(horizontal="left", vertical="center", wrap_text=False)
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = data_border
            if c == 1:
                cell.alignment = data_align_title
            else:
                cell.alignment = data_align_link

    # --- Column widths ---
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 45

    # --- Row heights ---
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 40

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # --- GUI-ready startup ---
    # Open LibreOffice Calc with the arxiv_papers.xlsx file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    # Open Chrome for the agent to download PDFs from ArXiv
    launch_gui('google-chrome', delay_sec=2.0)

    print('GUI_READY: launched LibreOffice Calc and Chrome with DISPLAY=:0')


create_initial()
