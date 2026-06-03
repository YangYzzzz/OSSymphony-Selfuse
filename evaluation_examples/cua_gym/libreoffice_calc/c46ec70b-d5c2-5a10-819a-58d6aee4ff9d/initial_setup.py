"""
Initial Setup: Data validation on cell C2 with custom formula for uppercase-only entries
Task ID: calc_nrv_069
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_069'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Inventory ---
    ws = wb.active
    ws.title = 'Inventory'

    # Headers
    headers = ['Item', 'Category', 'Code']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows — Items and Categories filled, Code column left empty
    # (agent task is to set up validation on C2, not fill codes)
    data = [
        ['Wireless Bluetooth Headphones', 'Electronics'],
        ['Organic Green Tea - 100 bags', 'Beverages'],
        ['Ergonomic Office Chair', 'Furniture'],
        ['Stainless Steel Water Bottle 750ml', 'Kitchen'],
        ['LED Desk Lamp with USB Port', 'Electronics'],
        ['Natural Bamboo Cutting Board', 'Kitchen'],
        ['Adjustable Standing Desk Converter', 'Furniture'],
        ['Premium Notebook - A5 Lined', 'Stationery'],
        ['Ceramic Pour-Over Coffee Dripper', 'Kitchen'],
        ['Mechanical Keyboard - Cherry MX', 'Electronics'],
        ['Yoga Mat - 6mm Thick', 'Fitness'],
        ['Cast Iron Skillet 12 inch', 'Kitchen'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
        # C column (Code) left empty — no value, no validation
        ws.cell(row=r, column=3).border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 14

    # --- Sheet 2: Categories ---
    ws2 = wb.create_sheet('Categories')
    cat_headers = ['Category', 'Department', 'Budget']
    for col, h in enumerate(cat_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    categories = [
        ['Electronics', 'Technology', 25000],
        ['Beverages', 'Food & Drink', 8000],
        ['Furniture', 'Operations', 45000],
        ['Kitchen', 'Food & Drink', 12000],
        ['Stationery', 'Office Supplies', 3500],
        ['Fitness', 'Wellness', 6000],
    ]
    for r, row_data in enumerate(categories, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
