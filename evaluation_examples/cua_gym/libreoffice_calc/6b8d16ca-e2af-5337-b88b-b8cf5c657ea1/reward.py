"""
Reward Script: Sales Pipeline Tracker Setup
Task ID: calc_sales_pipeline_stage_001
Domain: libreoffice_calc
Scoring:
  Component 1: Column H (Expected Revenue) has =E*G formulas for rows 2-101 (0.35 pts)
  Component 2: Conditional formatting on Stage column D2:D101 with color rules (0.35 pts)
  Component 3: Summary sheet column B has SUMIFS formulas for each stage (0.30 pts)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_sales_pipeline_stage_001'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # -------------------------------------------------------------------------
    # Component 1: Column H has Expected Revenue formulas (0.35 points)
    # Task requires: column H must contain formula =E*G for rows 2-101
    # Initial state: column H is empty (None)
    # Golden state: H2=E2*G2, H3=E3*G3, ..., H101=E101*G101
    # -------------------------------------------------------------------------
    try:
        if 'Pipeline' not in wb.sheetnames:
            print("FAIL: Component 1 — 'Pipeline' sheet not found")
        else:
            ws_pipeline = wb['Pipeline']

            # Count cells in column H (rows 2-101) that have E*G formulas
            formula_count = 0
            wrong_formula_count = 0
            empty_count = 0

            for r in range(2, 102):
                val = ws_pipeline.cell(row=r, column=8).value
                if val is None:
                    empty_count += 1
                elif isinstance(val, str):
                    # Check for formula pattern: =E<row>*G<row> or equivalent
                    # Normalize: remove spaces, uppercase
                    formula_norm = val.upper().replace(' ', '')
                    # Accept variations: =E2*G2 or =G2*E2
                    expected1 = f'=E{r}*G{r}'
                    expected2 = f'=G{r}*E{r}'
                    # Also accept PRODUCT or other equivalent forms
                    if (formula_norm == expected1.upper() or
                            formula_norm == expected2.upper()):
                        formula_count += 1
                    else:
                        # Check if it's any formula referencing both E and G columns
                        if 'E' in formula_norm and 'G' in formula_norm and formula_norm.startswith('='):
                            formula_count += 1
                        else:
                            wrong_formula_count += 1
                else:
                    # Numeric value — formula was computed but might be stored as value
                    # We require formula presence, not computed value
                    wrong_formula_count += 1

            if formula_count == 100:
                print(f"PASS: Component 1 — All 100 rows (H2:H101) have Expected Revenue formulas (0.35 pts)")
                total_score += 0.35
            elif formula_count >= 80:
                print(f"PARTIAL: Component 1 — {formula_count}/100 rows have correct formula, {empty_count} empty, {wrong_formula_count} wrong")
                total_score += 0.15
            else:
                print(f"FAIL: Component 1 — Only {formula_count}/100 rows have Expected Revenue formula, {empty_count} empty, {wrong_formula_count} wrong")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Conditional formatting on Stage column D2:D101 (0.35 points)
    # Task requires:
    #   green for 'Closed Won'
    #   yellow for 'Negotiation' and 'Proposal'
    #   red for 'Prospecting' and 'Qualification'
    # Initial state: no conditional formatting on Pipeline sheet
    # Golden state: 5 rules on D2:D101 with specific fill colors
    # -------------------------------------------------------------------------
    try:
        if 'Pipeline' not in wb.sheetnames:
            print("FAIL: Component 2 — 'Pipeline' sheet not found")
        else:
            ws_pipeline = wb['Pipeline']
            cf_rules = ws_pipeline.conditional_formatting

            # Collect all rules applied to the D column range
            d_column_rules = []
            for cf_range in cf_rules:
                # Check if this CF range covers D column cells
                range_str = str(cf_range)
                if 'D' in range_str.upper():
                    for rule in cf_rules[cf_range]:
                        d_column_rules.append(rule)

            if len(d_column_rules) == 0:
                print(f"FAIL: Component 2 — No conditional formatting rules found on column D")
            else:
                # Check for each required stage rule
                stages_found = {
                    'Closed Won': False,
                    'Negotiation': False,
                    'Proposal': False,
                    'Prospecting': False,
                    'Qualification': False
                }

                for rule in d_column_rules:
                    if hasattr(rule, 'formula') and rule.formula:
                        formula_str = str(rule.formula).upper()
                        for stage in stages_found:
                            if stage.upper() in formula_str:
                                stages_found[stage] = True

                stages_covered = sum(1 for v in stages_found.values() if v)
                print(f"  CF rules found on D column: {len(d_column_rules)}")
                print(f"  Stages covered: {stages_covered}/5 — {[s for s, v in stages_found.items() if v]}")

                # Award points based on coverage
                if stages_covered == 5:
                    # Check color correctness: green for Closed Won, yellow for Negotiation/Proposal,
                    # red for Prospecting/Qualification
                    colors_correct = 0
                    for rule in d_column_rules:
                        if not (hasattr(rule, 'formula') and rule.formula and hasattr(rule, 'dxf') and rule.dxf):
                            continue
                        formula_str = str(rule.formula).upper()
                        fill = rule.dxf.fill if rule.dxf else None
                        if fill is None:
                            continue
                        try:
                            color_rgb = fill.fgColor.rgb.upper()
                        except Exception:
                            continue

                        # Check green for Closed Won
                        if 'CLOSED WON' in formula_str:
                            # Accept various shades of green: starts with FF and has high G component
                            # FF92D050 is the golden file's green
                            # Accept any color where the R component is lower than G
                            # More flexible: check it's not red/yellow
                            r_val = int(color_rgb[2:4], 16) if len(color_rgb) >= 8 else 0
                            g_val = int(color_rgb[4:6], 16) if len(color_rgb) >= 8 else 0
                            b_val = int(color_rgb[6:8], 16) if len(color_rgb) >= 8 else 0
                            if g_val > r_val and g_val > 100:  # greenish
                                colors_correct += 1

                        # Check yellow for Negotiation and Proposal
                        elif 'NEGOTIATION' in formula_str or 'PROPOSAL' in formula_str:
                            # Yellow: high R and G, low B. FFFFEB9C is golden file's yellow
                            r_val = int(color_rgb[2:4], 16) if len(color_rgb) >= 8 else 0
                            g_val = int(color_rgb[4:6], 16) if len(color_rgb) >= 8 else 0
                            b_val = int(color_rgb[6:8], 16) if len(color_rgb) >= 8 else 0
                            if r_val > 200 and g_val > 150 and b_val < 200:  # yellowish
                                colors_correct += 1

                        # Check red for Prospecting and Qualification
                        elif 'PROSPECTING' in formula_str or 'QUALIFICATION' in formula_str:
                            # Red: high R, low G, low B. FFFFC7CE is golden file's pink/red
                            r_val = int(color_rgb[2:4], 16) if len(color_rgb) >= 8 else 0
                            g_val = int(color_rgb[4:6], 16) if len(color_rgb) >= 8 else 0
                            b_val = int(color_rgb[6:8], 16) if len(color_rgb) >= 8 else 0
                            if r_val > 200 and g_val < 230 and b_val < 230:  # reddish/pinkish
                                colors_correct += 1

                    print(f"  Color checks passed: {colors_correct}/5")

                    if colors_correct >= 4:
                        print(f"PASS: Component 2 — All 5 stage rules with correct colors (0.35 pts)")
                        total_score += 0.35
                    elif colors_correct >= 2:
                        print(f"PARTIAL: Component 2 — All stages covered but only {colors_correct}/5 colors correct (0.20 pts)")
                        total_score += 0.20
                    else:
                        print(f"PARTIAL: Component 2 — All stages covered but colors incorrect (0.15 pts)")
                        total_score += 0.15

                elif stages_covered >= 3:
                    print(f"PARTIAL: Component 2 — Only {stages_covered}/5 stages covered (0.15 pts)")
                    total_score += 0.15
                else:
                    print(f"FAIL: Component 2 — Only {stages_covered}/5 stages covered in CF rules")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Summary sheet SUMIFS formulas for each stage (0.30 points)
    # Task requires: Summary!B2:B7 must have SUMIFS formulas for each stage
    #   that sum Expected Revenue (Pipeline!H) filtered by stage (Pipeline!D)
    # Initial state: Summary!B2:B7 are all None
    # Golden state: SUMIFS(Pipeline!$H$2:$H$101,Pipeline!$D$2:$D$101,A2..A7)
    # -------------------------------------------------------------------------
    try:
        if 'Summary' not in wb.sheetnames:
            print("FAIL: Component 3 — 'Summary' sheet not found")
        else:
            ws_summary = wb['Summary']

            # Check B2:B7 for SUMIFS formulas
            sumifs_count = 0
            non_formula_count = 0

            for r in range(2, 8):
                val = ws_summary.cell(row=r, column=2).value
                if val is None:
                    non_formula_count += 1
                elif isinstance(val, str):
                    formula_norm = val.upper().replace(' ', '')
                    if formula_norm.startswith('=SUMIFS('):
                        sumifs_count += 1
                    elif formula_norm.startswith('=SUMIF('):
                        # SUMIF (single condition) is acceptable
                        sumifs_count += 1
                    else:
                        non_formula_count += 1
                else:
                    # Numeric value — not a formula
                    non_formula_count += 1

            if sumifs_count == 6:
                print(f"PASS: Component 3 — All 6 Summary rows (B2:B7) have SUMIFS formulas (0.30 pts)")
                total_score += 0.30
            elif sumifs_count >= 4:
                print(f"PARTIAL: Component 3 — {sumifs_count}/6 Summary rows have SUMIFS formulas (0.15 pts)")
                total_score += 0.15
            elif sumifs_count >= 1:
                print(f"PARTIAL: Component 3 — Only {sumifs_count}/6 Summary rows have SUMIFS formulas (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 3 — No SUMIFS formulas in Summary!B2:B7 (found {non_formula_count} empty/non-formula cells)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
