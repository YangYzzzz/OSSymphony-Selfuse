"""
Initial Setup: Sales Pipeline Tracker
Task ID: calc_sales_pipeline_stage_001
Domain: libreoffice_calc

Creates a Pipeline sheet with 100 deals and a Summary sheet with stage labels.
Column H (Expected Revenue) is intentionally empty.
No conditional formatting or SUMIFS formulas — those are the task actions.
"""

import os
import random
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_pipeline_stage_001'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    random.seed(42)  # reproducible

    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ #
    # Sheet 1: Pipeline
    # ------------------------------------------------------------------ #
    ws = wb.active
    ws.title = 'Pipeline'

    # Headers
    headers = ['Deal ID', 'Company', 'Sales Rep', 'Stage', 'Deal Value',
               'Close Date', 'Win Probability', 'Expected Revenue']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Realistic data pools
    companies = [
        'Apex Technologies', 'Blue Horizon Corp', 'Cascade Systems', 'DataBridge Inc',
        'Elevate Solutions', 'Frontier Analytics', 'GlobalNet Services', 'Harbor Digital',
        'Innovatech LLC', 'Jade Consulting', 'Keystone Software', 'Luminary Group',
        'Meridian Partners', 'NexGen Platforms', 'Orbit Dynamics', 'Peak Performance Co',
        'Quantum Leap Ltd', 'Riverside Enterprises', 'Summit Cloud', 'TechVault Corp',
        'Unified Systems', 'Vertex Industries', 'Wavefront Media', 'Xenon Analytics',
        'Yellowstone Digital', 'Zenith Solutions', 'ArcLight Technologies', 'BrightPath Inc',
        'Clearwater Systems', 'DawnBreaker Corp',
    ]

    sales_reps = [
        'Sarah Chen', 'Marcus Johnson', 'Priya Patel', 'Daniel Rivera', 'Emily Foster',
        'Kevin Park', 'Natalie Brooks', 'James Okafor', 'Lauren White', 'Tyler Rodriguez',
        'Amanda Singh', 'Christopher Lee',
    ]

    stages = [
        'Prospecting', 'Qualification', 'Proposal', 'Negotiation', 'Closed Won', 'Closed Lost',
    ]

    # Stage distribution weights (realistic pipeline shape)
    stage_weights = [0.20, 0.20, 0.20, 0.15, 0.15, 0.10]

    base_date = datetime.date(2026, 1, 1)

    for row in range(2, 102):  # 100 deals
        deal_id = f'DEAL-{2025000 + row - 1}'
        company = companies[(row - 2) % len(companies)]
        # Add variation to company names
        if (row - 2) >= len(companies):
            suffix_idx = (row - 2) // len(companies)
            suffixes = [' (West)', ' (East)', ' (North)', ' (South)', ' II']
            company = company + suffixes[(suffix_idx - 1) % len(suffixes)]
        rep = sales_reps[(row - 2) % len(sales_reps)]
        stage = random.choices(stages, weights=stage_weights, k=1)[0]
        deal_value = round(random.choice([
            5000, 8500, 12000, 15750, 22000, 28000, 35000, 42500,
            55000, 68000, 75000, 90000, 110000, 135000, 160000,
            185000, 210000, 245000, 280000, 315000, 350000,
        ]) + random.uniform(-500, 500), -2)
        days_ahead = random.randint(10, 180)
        close_date = base_date + datetime.timedelta(days=days_ahead)
        win_prob = {
            'Prospecting': round(random.uniform(0.05, 0.15), 2),
            'Qualification': round(random.uniform(0.15, 0.30), 2),
            'Proposal': round(random.uniform(0.30, 0.55), 2),
            'Negotiation': round(random.uniform(0.55, 0.80), 2),
            'Closed Won': 1.00,
            'Closed Lost': 0.00,
        }[stage]

        ws.cell(row=row, column=1, value=deal_id)
        ws.cell(row=row, column=2, value=company)
        ws.cell(row=row, column=3, value=rep)
        ws.cell(row=row, column=4, value=stage)
        ws.cell(row=row, column=5, value=deal_value)
        ws.cell(row=row, column=6, value=close_date)
        ws.cell(row=row, column=7, value=win_prob)
        # Column H (Expected Revenue) left EMPTY intentionally — task asks to add formula

    # Column widths for readability
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 20

    # ------------------------------------------------------------------ #
    # Sheet 2: Summary
    # ------------------------------------------------------------------ #
    ws2 = wb.create_sheet('Summary')

    ws2['A1'] = 'Stage'
    ws2['B1'] = 'Total Expected Revenue'
    ws2['A1'].font = Font(bold=True)
    ws2['B1'].font = Font(bold=True)

    # Stage labels in A2:A7
    stage_labels = ['Prospecting', 'Qualification', 'Proposal',
                    'Negotiation', 'Closed Won', 'Closed Lost']
    for r, label in enumerate(stage_labels, 2):
        ws2.cell(row=r, column=1, value=label)
    # Column B (Total Expected Revenue) left EMPTY — task asks to add SUMIFS formulas

    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 26

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Pipeline sheet: 100 deals, rows 2-101')
    print(f'  Summary sheet: Stage labels A2:A7, Column B empty')
    print(f'  Column H (Expected Revenue): empty')
    print(f'  No conditional formatting applied')
    print(f'  No SUMIFS formulas in Summary')

create_initial()
