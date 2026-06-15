"""
Reward Script: Pie chart with title, percentage labels, and exploded largest slice
Task ID: calc_gg3_002
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.25): PieChart exists on Budget sheet
  - Component 2 (0.25): Chart title is 'Annual Budget Allocation'
  - Component 3 (0.25): Percentage data labels enabled
  - Component 4 (0.25): Largest slice (Salaries, idx 0) exploded by ~10%
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_002'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice session."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl
    from openpyxl.chart import PieChart

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Budget sheet must exist
    if 'Budget' not in wb.sheetnames:
        print("FAIL: 'Budget' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Budget']

    # Component 1: A PieChart exists on the Budget sheet (0.25 points)
    try:
        charts = ws._charts
        pie_charts = [c for c in charts if isinstance(c, PieChart)]
        if len(pie_charts) >= 1:
            print(f"PASS: Component 1 — PieChart found on Budget sheet ({len(pie_charts)} pie chart(s)) (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — No PieChart found. Total charts: {len(charts)}, types: {[type(c).__name__ for c in charts]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Chart title is 'Annual Budget Allocation' (0.25 points)
    try:
        if pie_charts:
            chart = pie_charts[0]
            title_text = None
            if chart.title is not None:
                # Extract title text from the rich text structure
                try:
                    # Try direct string first
                    if isinstance(chart.title, str):
                        title_text = chart.title
                    elif hasattr(chart.title, 'tx') and chart.title.tx is not None:
                        rich = chart.title.tx.rich
                        if rich is not None:
                            parts = []
                            for p in rich.p:
                                for r in p.r:
                                    if r.t:
                                        parts.append(r.t)
                            title_text = ''.join(parts)
                except Exception:
                    title_text = str(chart.title)

            if title_text and title_text.strip() == 'Annual Budget Allocation':
                print(f"PASS: Component 2 — Chart title is 'Annual Budget Allocation' (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 2 — Expected title 'Annual Budget Allocation', found: '{title_text}'")
        else:
            print("FAIL: Component 2 — No PieChart to check title on")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Percentage data labels enabled (0.25 points)
    try:
        if pie_charts:
            chart = pie_charts[0]
            show_pct = False
            # Check chart-level dataLabels
            if hasattr(chart, 'dataLabels') and chart.dataLabels is not None:
                if chart.dataLabels.showPercent:
                    show_pct = True
            # Also check series-level dLbls
            if not show_pct:
                for s in chart.series:
                    if hasattr(s, 'dLbls') and s.dLbls is not None:
                        if s.dLbls.showPercent:
                            show_pct = True
                            break

            if show_pct:
                print(f"PASS: Component 3 — Percentage data labels enabled (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 3 — Percentage data labels not enabled (showPercent is not True)")
        else:
            print("FAIL: Component 3 — No PieChart to check data labels on")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Largest slice exploded by ~10% (0.25 points)
    # The largest value is Salaries (1,850,000) at index 0 in the data range A2:B6
    try:
        if pie_charts:
            chart = pie_charts[0]
            # Find the index of the largest value in the Budget data
            values = []
            for row_idx in range(2, 7):  # rows 2-6
                val = ws.cell(row=row_idx, column=2).value
                values.append(val if val is not None else 0)
            max_idx = values.index(max(values))  # should be 0 (Salaries)

            # Check if the data point at that index has explosion set
            exploded = False
            explosion_val = None
            for dp in chart.series[0].data_points:
                if dp.idx == max_idx and dp.explosion is not None:
                    explosion_val = dp.explosion
                    # Accept explosion in range 5-25 (approximately 10%)
                    if 5 <= dp.explosion <= 25:
                        exploded = True
                        break

            if exploded:
                print(f"PASS: Component 4 — Largest slice (idx {max_idx}) exploded by {explosion_val}% (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 4 — Largest slice (idx {max_idx}) not properly exploded. Explosion value: {explosion_val}")
        else:
            print("FAIL: Component 4 — No PieChart to check explosion on")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
