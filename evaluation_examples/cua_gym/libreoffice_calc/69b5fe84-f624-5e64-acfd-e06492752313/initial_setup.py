"""
Initial Setup: Create a budget spreadsheet with category and amount data for pie chart task.
Task ID: calc_gg3_002
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_002'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Budget Sheet ---
    ws = wb.active
    ws.title = 'Budget'

    # Headers
    ws.cell(row=1, column=1, value='Category')
    ws.cell(row=1, column=2, value='Amount')

    # Data rows - realistic budget amounts
    # Salaries is the largest category (typical for most organizations)
    data = [
        ['Salaries', 1850000],
        ['Marketing', 425000],
        ['R&D', 780000],
        ['Operations', 560000],
        ['Admin', 285000],
    ]

    for r, (category, amount) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=category)
        ws.cell(row=r, column=2, value=amount)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15

    # Format amounts as currency
    for r in range(2, 7):
        ws.cell(row=r, column=2).number_format = '#,##0'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
