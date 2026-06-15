"""
Reward Script: Quarterly Sales Performance Timesheet and KPI Summary
Task ID: calc_grs_016
Domain: libreoffice_calc
Scoring:
  C1: KPI Summary sheet exists (0.15)
  C2: Summary table with SUMIF formulas for revenue/units by rep (0.20)
  C3: Achievement % column with formula referencing revenue/target (0.10)
  C4: Summary sorted by total revenue descending (0.15)
  C5: Conditional formatting 3-color scale on achievement column (0.15)
  C6: Column chart on KPI Summary with 2 series (revenue vs target) (0.15)
  C7: Chart has title and axis labels (0.10)
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_016'


def persist_app_state(domain):
    """Best-effort save of any open LibreOffice document."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def compute_revenue_by_rep(ws_data, max_row):
    """Compute total revenue per sales rep from raw data."""
    from collections import defaultdict
    rev = defaultdict(int)
    units = defaultdict(int)
    target = defaultdict(int)
    for r in range(2, max_row + 1):
        rep = ws_data.cell(row=r, column=2).value
        if rep is None:
            continue
        u = ws_data.cell(row=r, column=5).value or 0
        rv = ws_data.cell(row=r, column=6).value or 0
        tg = ws_data.cell(row=r, column=7).value or 0
        rev[rep] += rv
        units[rep] += u
        target[rep] += tg
    return rev, units, target


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sales Data sheet must exist
    if 'Sales Data' not in wb.sheetnames:
        # Check if there's a Sheet1-like name with raw data
        has_raw = False
        for sn in wb.sheetnames:
            ws_check = wb[sn]
            if ws_check.cell(row=1, column=1).value == 'Date' and ws_check.cell(row=1, column=2).value == 'Sales Rep':
                has_raw = True
                break
        if not has_raw:
            print("CRITICAL: No raw data sheet found")
            print("REWARD: 0.0")
            return 0.0

    # Find the raw data sheet
    raw_sheet_name = None
    for sn in wb.sheetnames:
        ws_check = wb[sn]
        h1 = ws_check.cell(row=1, column=1).value
        h2 = ws_check.cell(row=1, column=2).value
        if h1 and 'Date' in str(h1) and h2 and 'Sales' in str(h2):
            raw_sheet_name = sn
            break
    if raw_sheet_name is None:
        print("CRITICAL: Cannot find raw data sheet")
        print("REWARD: 0.0")
        return 0.0

    ws_data = wb[raw_sheet_name]
    data_max_row = ws_data.max_row or 1

    # Compute expected values from raw data
    expected_rev, expected_units, expected_target = compute_revenue_by_rep(ws_data, data_max_row)
    expected_sorted = sorted(expected_rev.keys(), key=lambda x: expected_rev[x], reverse=True)

    # =========================================================================
    # Component 1: KPI Summary / summary sheet exists (0.15 points)
    # This is the KEY task-introduced change: initial file has only 1 sheet.
    # =========================================================================
    summary_ws = None
    summary_name = None
    try:
        # Look for a second sheet that has summary-like content
        for sn in wb.sheetnames:
            if sn == raw_sheet_name:
                continue
            ws_candidate = wb[sn]
            # Check if it has sales rep names and looks like a summary
            cell_a1 = ws_candidate.cell(row=1, column=1).value
            if cell_a1 is not None:
                summary_ws = ws_candidate
                summary_name = sn
                break

        if summary_ws is not None and len(wb.sheetnames) >= 2:
            print(f"PASS: Component 1 -- Summary sheet '{summary_name}' exists (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 -- No summary sheet found. Sheets: {wb.sheetnames}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    if summary_ws is None:
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # =========================================================================
    # Component 2: Summary table uses SUMIF formulas for revenue and units (0.20 pts)
    # Initial file has no summary sheet, so any SUMIF formulas are task-introduced.
    # =========================================================================
    try:
        sumif_count = 0
        total_sumif_expected = 0

        # Find which columns contain revenue and units SUMIF formulas
        # We need at least revenue SUMIF and units SUMIF for each rep
        summary_max_row = summary_ws.max_row or 1
        summary_max_col = summary_ws.max_column or 1

        for r in range(2, min(summary_max_row + 1, 20)):
            for c in range(1, min(summary_max_col + 1, 10)):
                val = summary_ws.cell(row=r, column=c).value
                if isinstance(val, str) and 'SUMIF' in val.upper():
                    sumif_count += 1

        # We expect at least 2 SUMIF per rep (revenue + units) for 6 reps = 12
        # But at minimum we need some SUMIF formulas
        if sumif_count >= 6:
            print(f"PASS: Component 2 -- Found {sumif_count} SUMIF formulas in summary (0.20 pts)")
            total_score += 0.20
        elif sumif_count >= 3:
            partial = 0.10
            print(f"PARTIAL: Component 2 -- Found {sumif_count} SUMIF formulas (partial {partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- Found only {sumif_count} SUMIF formulas, expected >= 6")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # =========================================================================
    # Component 3: Achievement % column with formula (0.10 pts)
    # Must reference revenue/target ratio. Not present in initial file.
    # =========================================================================
    try:
        ach_found = 0
        ach_col = None

        # Search header row for achievement-related column
        for c in range(1, min((summary_ws.max_column or 1) + 1, 15)):
            hdr = summary_ws.cell(row=1, column=c).value
            if hdr and ('achievement' in str(hdr).lower() or '%' in str(hdr)):
                ach_col = c
                break

        if ach_col is not None:
            for r in range(2, min((summary_ws.max_row or 1) + 1, 20)):
                val = summary_ws.cell(row=r, column=ach_col).value
                if isinstance(val, str) and '=' in val:
                    # It's a formula -- check it references a ratio or division
                    if '/' in val:
                        ach_found += 1

        if ach_found >= 3:
            print(f"PASS: Component 3 -- Achievement % formulas found ({ach_found} formulas) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 -- Achievement % column not found or missing formulas (found {ach_found})")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # =========================================================================
    # Component 4: Summary sorted by total revenue descending (0.15 pts)
    # Initial has no summary, so sort order is task-introduced.
    # We verify the rep order matches expected descending revenue sort.
    # =========================================================================
    try:
        actual_reps = []
        for r in range(2, min((summary_ws.max_row or 1) + 1, 20)):
            rep = summary_ws.cell(row=r, column=1).value
            if rep and isinstance(rep, str) and rep.strip():
                actual_reps.append(rep.strip())

        if len(actual_reps) == 0:
            print("FAIL: Component 4 -- No sales reps found in summary")
        else:
            # Check if the order matches expected descending revenue
            is_sorted = True
            matched = 0
            for i, exp_rep in enumerate(expected_sorted):
                if i < len(actual_reps) and actual_reps[i].strip() == exp_rep.strip():
                    matched += 1
                else:
                    is_sorted = False

            if is_sorted and matched == len(expected_sorted):
                print(f"PASS: Component 4 -- Summary correctly sorted by revenue descending (0.15 pts)")
                total_score += 0.15
            elif matched >= len(expected_sorted) - 1:
                # Almost correct sort, partial credit
                partial = 0.08
                print(f"PARTIAL: Component 4 -- Sort mostly correct ({matched}/{len(expected_sorted)} in order) ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 4 -- Sort order incorrect. Expected: {expected_sorted}, Got: {actual_reps}")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # =========================================================================
    # Component 5: Conditional formatting with 3-color scale on achievement column (0.15 pts)
    # Initial file has no conditional formatting at all.
    # =========================================================================
    try:
        cf_found = False
        cf_is_color_scale = False
        cf_has_3_colors = False

        for cf in summary_ws.conditional_formatting:
            for rule in cf.rules:
                if rule.type == 'colorScale' and rule.colorScale:
                    cf_found = True
                    cf_is_color_scale = True
                    num_colors = len(rule.colorScale.color) if rule.colorScale.color else 0
                    if num_colors >= 3:
                        cf_has_3_colors = True

        if cf_has_3_colors:
            print(f"PASS: Component 5 -- 3-color scale conditional formatting found (0.15 pts)")
            total_score += 0.15
        elif cf_is_color_scale:
            partial = 0.08
            print(f"PARTIAL: Component 5 -- Color scale found but not 3 colors ({partial} pts)")
            total_score += partial
        elif cf_found:
            partial = 0.05
            print(f"PARTIAL: Component 5 -- Some CF found but not color scale ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 -- No conditional formatting found on summary sheet")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    # =========================================================================
    # Component 6: Column chart on summary sheet with 2 series (0.15 pts)
    # Initial file has 0 charts. This is a task-introduced change.
    # =========================================================================
    try:
        charts = summary_ws._charts
        if len(charts) >= 1:
            chart = charts[0]
            is_col = (chart.type == 'col' or chart.type == 'bar')
            has_2_series = len(chart.series) >= 2

            if is_col and has_2_series:
                print(f"PASS: Component 6 -- Column chart with {len(chart.series)} series found (0.15 pts)")
                total_score += 0.15
            elif is_col:
                partial = 0.08
                print(f"PARTIAL: Component 6 -- Column chart found but only {len(chart.series)} series ({partial} pts)")
                total_score += partial
            elif has_2_series:
                partial = 0.08
                print(f"PARTIAL: Component 6 -- Chart with 2 series found but type is '{chart.type}' not 'col' ({partial} pts)")
                total_score += partial
            else:
                partial = 0.05
                print(f"PARTIAL: Component 6 -- A chart exists but type='{chart.type}', series={len(chart.series)} ({partial} pts)")
                total_score += partial
        else:
            print(f"FAIL: Component 6 -- No charts found on summary sheet")
    except Exception as e:
        print(f"ERROR: Component 6 -- {e}")

    # =========================================================================
    # Component 7: Chart has title and axis labels (0.10 pts)
    # Initial file has no charts, so any chart properties are task-introduced.
    # =========================================================================
    try:
        charts = summary_ws._charts
        if len(charts) >= 1:
            chart = charts[0]

            has_title = False
            has_y_axis = False
            has_x_axis = False

            # Check chart title
            if chart.title is not None:
                # Extract title text
                try:
                    title_text = chart.title.tx.rich.paragraphs[0].runs[0].t
                    if title_text and len(title_text.strip()) > 0:
                        has_title = True
                        print(f"  Chart title: '{title_text}'")
                except Exception:
                    # Title object exists but text extraction failed - still counts
                    has_title = True

            # Check y-axis title
            if chart.y_axis.title is not None:
                try:
                    y_text = chart.y_axis.title.tx.rich.paragraphs[0].runs[0].t
                    if y_text and len(y_text.strip()) > 0:
                        has_y_axis = True
                        print(f"  Y-axis title: '{y_text}'")
                except Exception:
                    has_y_axis = True

            # Check x-axis title
            if chart.x_axis.title is not None:
                try:
                    x_text = chart.x_axis.title.tx.rich.paragraphs[0].runs[0].t
                    if x_text and len(x_text.strip()) > 0:
                        has_x_axis = True
                        print(f"  X-axis title: '{x_text}'")
                except Exception:
                    has_x_axis = True

            labels_found = sum([has_title, has_y_axis, has_x_axis])
            if labels_found == 3:
                print(f"PASS: Component 7 -- Chart has title and both axis labels (0.10 pts)")
                total_score += 0.10
            elif labels_found >= 2:
                partial = 0.07
                print(f"PARTIAL: Component 7 -- Chart has {labels_found}/3 labels ({partial} pts)")
                total_score += partial
            elif labels_found >= 1:
                partial = 0.03
                print(f"PARTIAL: Component 7 -- Chart has {labels_found}/3 labels ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 7 -- Chart has no title or axis labels")
        else:
            print(f"FAIL: Component 7 -- No charts found, cannot check labels")
    except Exception as e:
        print(f"ERROR: Component 7 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
