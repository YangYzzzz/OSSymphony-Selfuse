"""
Initial Setup: Quarterly sales performance timesheet with raw data
Task ID: calc_grs_016
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_016'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: Sales Data (raw transaction data) ---
    ws1 = wb.active
    ws1.title = 'Sales Data'

    headers = ['Date', 'Sales Rep', 'Region', 'Product Category', 'Units Sold', 'Revenue', 'Target']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # Q3 2025 transaction data from CRM - 6 sales reps, multiple transactions each
    data = [
        [date(2025, 7, 3),  'Sarah Chen',      'West',      'Enterprise Software', 12,  84000,  75000],
        [date(2025, 7, 8),  'Marcus Johnson',   'East',      'Cloud Services',      25,  62500,  70000],
        [date(2025, 7, 12), 'Priya Patel',      'Central',   'Hardware',            18,  54000,  50000],
        [date(2025, 7, 15), 'David Kim',        'West',      'Enterprise Software', 8,   56000,  60000],
        [date(2025, 7, 19), 'Rachel Torres',    'Southeast', 'Cloud Services',      30,  75000,  65000],
        [date(2025, 7, 22), 'James O\'Brien',   'Northeast', 'Hardware',            15,  45000,  55000],
        [date(2025, 7, 28), 'Sarah Chen',       'West',      'Cloud Services',      20,  50000,  45000],
        [date(2025, 8, 2),  'Marcus Johnson',   'East',      'Enterprise Software', 10,  70000,  72000],
        [date(2025, 8, 5),  'Priya Patel',      'Central',   'Cloud Services',      22,  55000,  48000],
        [date(2025, 8, 11), 'David Kim',        'West',      'Hardware',            14,  42000,  40000],
        [date(2025, 8, 14), 'Rachel Torres',    'Southeast', 'Enterprise Software', 9,   63000,  58000],
        [date(2025, 8, 18), 'James O\'Brien',   'Northeast', 'Cloud Services',      28,  70000,  62000],
        [date(2025, 8, 22), 'Sarah Chen',       'West',      'Hardware',            16,  48000,  50000],
        [date(2025, 8, 27), 'Marcus Johnson',   'East',      'Hardware',            20,  60000,  55000],
        [date(2025, 9, 1),  'Priya Patel',      'Central',   'Enterprise Software', 11,  77000,  70000],
        [date(2025, 9, 5),  'David Kim',        'West',      'Cloud Services',      24,  60000,  58000],
        [date(2025, 9, 9),  'Rachel Torres',    'Southeast', 'Hardware',            17,  51000,  48000],
        [date(2025, 9, 12), 'James O\'Brien',   'Northeast', 'Enterprise Software', 7,   49000,  52000],
        [date(2025, 9, 16), 'Sarah Chen',       'West',      'Enterprise Software', 14,  98000,  85000],
        [date(2025, 9, 19), 'Marcus Johnson',   'East',      'Cloud Services',      32,  80000,  75000],
        [date(2025, 9, 23), 'Priya Patel',      'Central',   'Hardware',            19,  57000,  52000],
        [date(2025, 9, 25), 'David Kim',        'West',      'Enterprise Software', 13,  91000,  80000],
        [date(2025, 9, 28), 'Rachel Torres',    'Southeast', 'Cloud Services',      26,  65000,  60000],
        [date(2025, 9, 30), 'James O\'Brien',   'Northeast', 'Hardware',            21,  63000,  58000],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Format date column
    for r in range(2, len(data) + 2):
        ws1.cell(row=r, column=1).number_format = 'yyyy-mm-dd'

    # Format currency columns (Revenue and Target)
    for r in range(2, len(data) + 2):
        ws1.cell(row=r, column=6).number_format = '$#,##0'
        ws1.cell(row=r, column=7).number_format = '$#,##0'

    # Set reasonable column widths
    ws1.column_dimensions['A'].width = 14
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 12
    ws1.column_dimensions['D'].width = 22
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 12
    ws1.column_dimensions['G'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
