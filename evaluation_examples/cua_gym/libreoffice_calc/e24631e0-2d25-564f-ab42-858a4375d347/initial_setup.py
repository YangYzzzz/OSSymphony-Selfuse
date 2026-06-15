"""
Initial Setup: Build a client billing summary with hours by project, rate tiers,
and formatted invoice-ready output.
Task ID: calc_gpm_066
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_066'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

NAVY_ARGB = "FF000050"       # dark navy (0,0,80)
WHITE_ARGB = "FFFFFFFF"
LIGHT_GRAY_ARGB = "FFD9E1F2"  # light blue-gray for tier header


def launch_gui(command: str, delay_sec: float = 1.0):
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Billing"

    thin = Side(style="thin", color="000000")
    all_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ---- Title: Merge A1:G1 ----
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "Monthly Client Billing Summary - March 2026"
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color=NAVY_ARGB, end_color=NAVY_ARGB, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # ---- Rate Tier Reference Table I2:J5 ----
    tier_headers = [("Tier", "Rate")]
    tier_data = [("Standard", 150), ("Premium", 225), ("Emergency", 350)]

    ws["I2"] = "Tier"
    ws["J2"] = "Rate"
    for c in ["I2", "J2"]:
        ws[c].font = Font(bold=True, color="FFFFFF")
        ws[c].fill = PatternFill(start_color=NAVY_ARGB, end_color=NAVY_ARGB, fill_type="solid")
        ws[c].alignment = Alignment(horizontal="center")
        ws[c].border = all_border

    for i, (tier, rate) in enumerate(tier_data, 3):
        ws.cell(row=i, column=9, value=tier).border = all_border
        ws.cell(row=i, column=10, value=rate).border = all_border
        ws.cell(row=i, column=10).number_format = '$#,##0.00'

    # ---- Main Table Headers Row 3 ----
    headers = ["Client", "Project", "Tier", "Hours", "Rate", "Amount", "Status"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=NAVY_ARGB, end_color=NAVY_ARGB, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = all_border

    # ---- 14 Billing Entries (rows 4-17) ----
    billing_data = [
        # Client, Project, Tier, Hours, Status
        ("Meridian Corp", "Website Redesign", "Premium", 24.5, "Billed"),
        ("Meridian Corp", "SEO Optimization", "Standard", 18.0, "Billed"),
        ("Meridian Corp", "Mobile App Phase 1", "Emergency", 8.0, "Pending"),
        ("Apex Industries", "ERP Integration", "Premium", 32.0, "Billed"),
        ("Apex Industries", "Data Migration", "Standard", 45.5, "Pending"),
        ("Apex Industries", "Security Audit", "Emergency", 12.0, "Disputed"),
        ("Horizon Labs", "Dashboard Analytics", "Premium", 28.0, "Billed"),
        ("Horizon Labs", "API Development", "Standard", 36.0, "Pending"),
        ("Horizon Labs", "Load Testing", "Standard", 14.5, "Billed"),
        ("Silverline Finance", "Portfolio Tracker", "Premium", 40.0, "Billed"),
        ("Silverline Finance", "Compliance Report", "Emergency", 6.5, "Disputed"),
        ("Silverline Finance", "Client Portal", "Standard", 22.0, "Pending"),
        ("NovaTech Solutions", "Cloud Migration", "Premium", 35.0, "Billed"),
        ("NovaTech Solutions", "DevOps Pipeline", "Standard", 19.0, "Pending"),
    ]

    for r, (client, project, tier, hours, status) in enumerate(billing_data, 4):
        ws.cell(row=r, column=1, value=client).border = all_border
        ws.cell(row=r, column=2, value=project).border = all_border
        ws.cell(row=r, column=3, value=tier).border = all_border
        c_hours = ws.cell(row=r, column=4, value=hours)
        c_hours.number_format = '0.0'
        c_hours.border = all_border
        # Rate column (E) - leave EMPTY for the task (agent must add VLOOKUP)
        c_rate = ws.cell(row=r, column=5)
        c_rate.number_format = '$#,##0.00'
        c_rate.border = all_border
        # Amount column (F) - leave EMPTY (agent must add =D*E)
        c_amount = ws.cell(row=r, column=6)
        c_amount.number_format = '$#,##0.00'
        c_amount.border = all_border
        ws.cell(row=r, column=7, value=status).border = all_border

    # ---- Data Validation: Tier dropdown C4:C17 ----
    dv_tier = DataValidation(
        type="list",
        formula1='"Standard,Premium,Emergency"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_tier.error = "Invalid tier"
    dv_tier.errorTitle = "Error"
    dv_tier.add("C4:C17")
    ws.add_data_validation(dv_tier)

    # ---- Data Validation: Status dropdown G4:G17 ----
    dv_status = DataValidation(
        type="list",
        formula1='"Billed,Pending,Disputed"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_status.error = "Invalid status"
    dv_status.errorTitle = "Error"
    dv_status.add("G4:G17")
    ws.add_data_validation(dv_status)

    # ---- Column widths ----
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
