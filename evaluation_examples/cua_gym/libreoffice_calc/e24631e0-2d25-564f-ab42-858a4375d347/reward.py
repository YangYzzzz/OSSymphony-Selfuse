"""
Reward Script: Client billing summary with hours, rate tiers, VLOOKUP, SUMIFS, and formatting
Task ID: calc_gpm_066
Domain: libreoffice_calc
Scoring:
  Component 1: VLOOKUP formulas in E4:E17 (0.25)
  Component 2: Amount formulas in F4:F17 (0.20)
  Component 3: Conditional formatting on G4:G17 (0.15)
  Component 4: Client Summary section rows 19-24 with SUMIFS (0.20)
  Component 5: Grand Total row 26 with SUM + double border (0.20)
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_066'


def persist_app_state(domain):
    """Save any unsaved GUI edits before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Billing' not in wb.sheetnames:
        print("CRITICAL: 'Billing' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Billing']

    # Component 1: VLOOKUP formulas in E4:E17 (0.25 points)
    # Initial env has EMPTY E4:E17. Golden has =VLOOKUP(C,tier_table,2,FALSE).
    try:
        vlookup_count = 0
        for r in range(4, 18):
            val = ws.cell(row=r, column=5).value  # col E
            if val is not None and isinstance(val, str) and 'VLOOKUP' in val.upper():
                vlookup_count += 1
        if vlookup_count == 14:
            print(f"PASS: Component 1 — All 14 VLOOKUP formulas found in E4:E17 (0.25 pts)")
            total_score += 0.25
        elif vlookup_count >= 7:
            partial = round(0.25 * vlookup_count / 14, 2)
            print(f"PARTIAL: Component 1 — {vlookup_count}/14 VLOOKUP formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Expected 14 VLOOKUP formulas in E4:E17, found {vlookup_count}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Amount formulas (=D*E) in F4:F17 (0.20 points)
    # Initial env has EMPTY F4:F17. Golden has multiplication formulas.
    try:
        amount_count = 0
        for r in range(4, 18):
            val = ws.cell(row=r, column=6).value  # col F
            if val is not None and isinstance(val, str):
                # Check for multiplication formula referencing D and E columns
                v_upper = val.upper().replace(' ', '')
                if ('D' in v_upper and 'E' in v_upper and '=' in v_upper) or ('*' in v_upper and '=' in v_upper):
                    amount_count += 1
        if amount_count == 14:
            print(f"PASS: Component 2 — All 14 amount formulas found in F4:F17 (0.20 pts)")
            total_score += 0.20
        elif amount_count >= 7:
            partial = round(0.20 * amount_count / 14, 2)
            print(f"PARTIAL: Component 2 — {amount_count}/14 amount formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Expected 14 amount formulas in F4:F17, found {amount_count}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Conditional formatting on G4:G17 (0.15 points)
    # Initial env has NO conditional formatting. Golden has 3 rules on G4:G17.
    try:
        cf_on_g = 0
        has_billed_rule = False
        has_pending_rule = False
        has_disputed_rule = False
        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            # Check if G column is in the conditional formatting range
            if 'G' in cf_range:
                for rule in cf.rules:
                    if rule.type == 'cellIs':
                        formulas = [str(f).strip('"').strip("'") for f in rule.formula]
                        for f in formulas:
                            if 'Billed' in f:
                                has_billed_rule = True
                            elif 'Pending' in f:
                                has_pending_rule = True
                            elif 'Disputed' in f:
                                has_disputed_rule = True

        rules_found = sum([has_billed_rule, has_pending_rule, has_disputed_rule])
        if rules_found == 3:
            print(f"PASS: Component 3 — All 3 conditional formatting rules found on G column (0.15 pts)")
            total_score += 0.15
        elif rules_found >= 1:
            partial = round(0.15 * rules_found / 3, 2)
            print(f"PARTIAL: Component 3 — {rules_found}/3 conditional formatting rules found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No conditional formatting rules found on G column")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Client Summary section rows 19-24 with SUMIFS (0.20 points)
    # Initial env has EMPTY rows 19-26. Golden has Client Summary header + 5 client rows with SUMIFS.
    try:
        sub_score = 0.0

        # 4a: Row 19 header "Client Summary" (0.04 pts)
        a19_val = ws['A19'].value
        if a19_val is not None and 'Client Summary' in str(a19_val):
            print(f"  4a PASS: A19 = {a19_val!r}")
            sub_score += 0.04
        else:
            print(f"  4a FAIL: A19 = {a19_val!r}, expected 'Client Summary'")

        # 4b: Client names in A20:A24 (0.04 pts)
        expected_clients = ['Meridian Corp', 'Apex Industries', 'Horizon Labs', 'Silverline Finance', 'NovaTech Solutions']
        found_clients = 0
        for r in range(20, 25):
            val = ws.cell(row=r, column=1).value
            if val is not None and str(val).strip() in expected_clients:
                found_clients += 1
        if found_clients == 5:
            print(f"  4b PASS: All 5 client names found in A20:A24")
            sub_score += 0.04
        else:
            print(f"  4b FAIL: {found_clients}/5 client names found")

        # 4c: SUMIFS formulas in summary rows (0.08 pts)
        # Check columns C and D for rows 20-24
        sumifs_count = 0
        for r in range(20, 25):
            for c in [3, 4]:  # columns C and D
                val = ws.cell(row=r, column=c).value
                if val is not None and isinstance(val, str) and 'SUMIF' in val.upper():
                    sumifs_count += 1
        # Expect 10 SUMIFS formulas (5 clients x 2 columns)
        if sumifs_count >= 8:
            print(f"  4c PASS: {sumifs_count} SUMIFS formulas found in summary ({0.08} pts)")
            sub_score += 0.08
        elif sumifs_count >= 4:
            partial = round(0.08 * sumifs_count / 10, 2)
            print(f"  4c PARTIAL: {sumifs_count}/10 SUMIFS formulas found ({partial} pts)")
            sub_score += partial
        else:
            print(f"  4c FAIL: Only {sumifs_count} SUMIFS formulas found in summary rows")

        # 4d: Merged cells in summary A20:B20 through A24:B24 (0.04 pts)
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        summary_merges = 0
        for r in range(20, 25):
            expected_merge = f'A{r}:B{r}'
            if expected_merge in merged_ranges:
                summary_merges += 1
        if summary_merges >= 4:
            print(f"  4d PASS: {summary_merges}/5 summary row merges found")
            sub_score += 0.04
        else:
            print(f"  4d FAIL: {summary_merges}/5 summary row merges found")

        if sub_score > 0:
            print(f"PASS: Component 4 — Client Summary section ({sub_score} pts)")
        else:
            print(f"FAIL: Component 4 — Client Summary section missing")
        total_score += sub_score
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Grand Total row 26 with SUM formula + double border (0.20 points)
    # Initial env has EMPTY row 26. Golden has Grand Total with SUM and double border.
    try:
        sub_score = 0.0

        # 5a: "Grand Total" label in A26 (0.05 pts)
        a26_val = ws['A26'].value
        if a26_val is not None and 'Grand Total' in str(a26_val):
            print(f"  5a PASS: A26 = {a26_val!r}")
            sub_score += 0.05
        else:
            print(f"  5a FAIL: A26 = {a26_val!r}, expected 'Grand Total'")

        # 5b: A26 bold (0.03 pts)
        if ws['A26'].font.bold:
            print(f"  5b PASS: A26 is bold")
            sub_score += 0.03
        else:
            print(f"  5b FAIL: A26 not bold")

        # 5c: SUM formula in grand total row (0.07 pts)
        sum_found = False
        for c in range(3, 7):  # columns C through F
            val = ws.cell(row=26, column=c).value
            if val is not None and isinstance(val, str) and 'SUM' in val.upper():
                sum_found = True
                break
        if sum_found:
            print(f"  5c PASS: SUM formula found in row 26")
            sub_score += 0.07
        else:
            print(f"  5c FAIL: No SUM formula found in row 26")

        # 5d: Double border on row 26 (0.05 pts)
        has_double_border = False
        for c in range(1, 5):
            border_style = ws.cell(row=26, column=c).border.bottom.style
            if border_style == 'double':
                has_double_border = True
                break
        if has_double_border:
            print(f"  5d PASS: Double border found on row 26")
            sub_score += 0.05
        else:
            print(f"  5d FAIL: No double border on row 26")

        if sub_score > 0:
            print(f"PASS: Component 5 — Grand Total row ({sub_score} pts)")
        else:
            print(f"FAIL: Component 5 — Grand Total row missing")
        total_score += sub_score
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
