"""
reward.py for calc_gg1_039
Verifies: thick blue bottom border on A1:Z1 in 'Report' sheet,
no unwanted borders (top/left/right) on row 1, no borders on row 2.
"""

import os
import traceback

def main():
    score = 0.0
    filepath = "/home/user/calc_gg1_039.xlsx"

    # Gate: file must exist
    if not os.path.exists(filepath):
        print("File not found:", filepath)
        print(f"REWARD: 0.0")
        return

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("openpyxl not available")
        print(f"REWARD: 0.0")
        return

    wb = None
    try:
        wb = load_workbook(filepath)
    except Exception as e:
        print("Failed to load workbook:", e)
        print(f"REWARD: 0.0")
        return

    # Precondition: File has 'Report' sheet (gate, no points)
    try:
        if "Report" not in wb.sheetnames:
            print("FAIL: 'Report' sheet not found. Sheets:", wb.sheetnames)
            print(f"REWARD: 0.0")
            return
        print("Gate PASS: 'Report' sheet found")
    except Exception as e:
        print("Gate ERROR:", e)
        print(f"REWARD: 0.0")
        return

    ws = wb["Report"]

    # Component 1: Bottom border present on A1:Z1, medium or thick (0.5 pts)
    comp1 = 0.0
    cells_with_bottom = 0
    try:
        thick_styles = {"medium", "thick", "mediumDashed", "mediumDashDot", "mediumDashDotDot"}
        for col_idx in range(1, 27):  # A=1 to Z=26
            cell = ws.cell(row=1, column=col_idx)
            bottom = cell.border.bottom
            if bottom and bottom.style and bottom.style != "none":
                # Accept medium or thick (not thin/hair)
                if bottom.style in thick_styles:
                    cells_with_bottom += 1
                else:
                    print(f"  Col {col_idx}: bottom border style='{bottom.style}' (not medium/thick)")
        comp1 = 0.5 * (cells_with_bottom / 26)
        print(f"Component 1: {cells_with_bottom}/26 cells have medium/thick bottom border -> {comp1:.4f}")
    except Exception as e:
        print("Component 1 ERROR:", e)
        traceback.print_exc()

    # Component 2: Border color is blue (0.3 pts)
    comp2 = 0.0
    try:
        cells_with_blue = 0
        for col_idx in range(1, 27):
            cell = ws.cell(row=1, column=col_idx)
            bottom = cell.border.bottom
            if bottom and bottom.style and bottom.style != "none":
                color = bottom.color
                if color is not None:
                    rgb = None
                    try:
                        rgb = color.rgb
                    except Exception:
                        pass
                    if rgb and isinstance(rgb, str):
                        # ARGB format: e.g. 'FF0000FF' or '000000FF'
                        # Extract the RGB portion (last 6 chars)
                        rgb_hex = rgb[-6:].upper()
                        # Check if blue: high B value, low R and G
                        r = int(rgb_hex[0:2], 16)
                        g = int(rgb_hex[2:4], 16)
                        b = int(rgb_hex[4:6], 16)
                        # Blue means B is dominant
                        if b >= 200 and r <= 100 and g <= 100:
                            cells_with_blue += 1
                        else:
                            print(f"  Col {col_idx}: border color RGB=({r},{g},{b}) - not blue enough")
                    else:
                        print(f"  Col {col_idx}: border color rgb is None or non-string: {rgb}")
                else:
                    print(f"  Col {col_idx}: border color object is None")
        comp2 = 0.3 * (cells_with_blue / 26)
        print(f"Component 2: {cells_with_blue}/26 cells have blue bottom border -> {comp2:.4f}")
    except Exception as e:
        print("Component 2 ERROR:", e)
        traceback.print_exc()

    # Component 3: No unwanted borders (0.2 pts)
    # Only award if there IS a bottom border (comp1 > 0), otherwise 0
    comp3 = 0.0
    try:
        if cells_with_bottom == 0:
            print("Component 3: skipped (no bottom borders found, nothing to check)")
            comp3 = 0.0
        else:
            total_checks = 0
            violations = 0

            # Check row 1 cells for unwanted top/left/right borders
            for col_idx in range(1, 27):
                cell = ws.cell(row=1, column=col_idx)
                for side_name in ["top", "left", "right"]:
                    total_checks += 1
                    side = getattr(cell.border, side_name, None)
                    if side and side.style and side.style != "none":
                        violations += 1
                        if violations <= 5:
                            print(f"  Violation: row1 col{col_idx} has {side_name} border (style={side.style})")

            # Check row 2 cells for borders (should have none added)
            for col_idx in range(1, 27):
                cell = ws.cell(row=2, column=col_idx)
                for side_name in ["top", "bottom", "left", "right"]:
                    total_checks += 1
                    side = getattr(cell.border, side_name, None)
                    if side and side.style and side.style != "none":
                        violations += 1
                        if violations <= 10:
                            print(f"  Violation: row2 col{col_idx} has {side_name} border (style={side.style})")

            if total_checks > 0:
                clean_ratio = max(0, 1.0 - (violations / total_checks))
                comp3 = 0.2 * clean_ratio
            else:
                comp3 = 0.2

            print(f"Component 3: {violations} violations out of {total_checks} checks -> {comp3:.4f}")
    except Exception as e:
        print("Component 3 ERROR:", e)
        traceback.print_exc()

    score = round(comp1 + comp2 + comp3, 2)
    # Clamp to [0.0, 1.0]
    score = max(0.0, min(1.0, score))
    print(f"REWARD: {score}")


if __name__ == "__main__":
    main()
