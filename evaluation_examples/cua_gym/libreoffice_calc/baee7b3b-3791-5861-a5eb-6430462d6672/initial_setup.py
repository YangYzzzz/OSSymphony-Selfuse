import subprocess as _sp
_sp.check_call(["pip3", "install", "openpyxl"], stdout=_sp.DEVNULL, stderr=_sp.DEVNULL)

import openpyxl
from openpyxl.styles import Font, Alignment
import os
import subprocess
import time

output_path = "/home/user/calc_gg1_039.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Report"

# Header row A1:Z1 - 26 realistic department report columns
headers = [
    "Employee ID",        # A
    "First Name",         # B
    "Last Name",          # C
    "Department",         # D
    "Job Title",          # E
    "Hire Date",          # F
    "Years of Service",   # G
    "Salary",             # H
    "Bonus",              # I
    "Total Compensation", # J
    "Performance Rating", # K
    "Manager",            # L
    "Office Location",    # M
    "Phone Extension",    # N
    "Email",              # O
    "Training Hours",     # P
    "Certifications",     # Q
    "Projects Completed", # R
    "Sick Days Used",     # S
    "Vacation Days Used", # T
    "Overtime Hours",     # U
    "Travel Expenses",    # V
    "Equipment Cost",     # W
    "Software Licenses",  # X
    "Notes",              # Y
    "Status",             # Z
]

bold_font = Font(bold=True)
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = bold_font

# Data rows - 25 rows of realistic department report data
data_rows = [
    ["EMP001", "Alice", "Johnson", "Engineering", "Software Engineer", "2019-03-15", 7, 95000, 12000, 107000, 4.5, "Robert Chen", "New York", 1201, "alice.johnson@company.com", 40, 3, 12, 2, 10, 15, 2500, 1800, 3, "", "Active"],
    ["EMP002", "Brian", "Williams", "Marketing", "Marketing Manager", "2017-06-01", 9, 88000, 9500, 97500, 4.2, "Sarah Lee", "Chicago", 1302, "brian.williams@company.com", 24, 2, 8, 4, 12, 5, 4200, 900, 2, "", "Active"],
    ["EMP003", "Catherine", "Davis", "Finance", "Financial Analyst", "2020-01-10", 6, 82000, 8000, 90000, 4.0, "James Park", "New York", 1105, "catherine.davis@company.com", 32, 4, 10, 1, 8, 20, 1500, 1200, 4, "", "Active"],
    ["EMP004", "David", "Martinez", "Engineering", "Senior Developer", "2016-09-20", 10, 115000, 18000, 133000, 4.8, "Robert Chen", "San Francisco", 1203, "david.martinez@company.com", 50, 5, 18, 3, 15, 25, 3800, 2200, 5, "", "Active"],
    ["EMP005", "Emily", "Brown", "Human Resources", "HR Specialist", "2021-04-05", 5, 68000, 5500, 73500, 3.8, "Linda Torres", "Chicago", 1401, "emily.brown@company.com", 28, 2, 6, 5, 7, 0, 800, 600, 2, "", "Active"],
    ["EMP006", "Frank", "Garcia", "Sales", "Sales Representative", "2018-11-12", 8, 72000, 15000, 87000, 4.3, "Mark Wilson", "Dallas", 1501, "frank.garcia@company.com", 18, 1, 14, 2, 9, 10, 6500, 700, 1, "Top performer Q3", "Active"],
    ["EMP007", "Grace", "Miller", "Engineering", "QA Engineer", "2022-02-28", 4, 78000, 7000, 85000, 3.9, "Robert Chen", "New York", 1204, "grace.miller@company.com", 36, 3, 9, 6, 5, 12, 1200, 1500, 3, "", "Active"],
    ["EMP008", "Henry", "Wilson", "Finance", "Senior Accountant", "2015-07-14", 11, 98000, 11000, 109000, 4.1, "James Park", "New York", 1106, "henry.wilson@company.com", 20, 6, 15, 3, 14, 8, 2000, 1000, 4, "", "Active"],
    ["EMP009", "Irene", "Taylor", "Marketing", "Content Strategist", "2020-08-03", 6, 75000, 7500, 82500, 4.0, "Sarah Lee", "Chicago", 1303, "irene.taylor@company.com", 30, 2, 7, 1, 11, 3, 1800, 800, 2, "", "Active"],
    ["EMP010", "Jack", "Anderson", "Engineering", "DevOps Engineer", "2019-12-01", 7, 105000, 14000, 119000, 4.6, "Robert Chen", "San Francisco", 1205, "jack.anderson@company.com", 45, 7, 16, 2, 13, 30, 2200, 2500, 6, "", "Active"],
    ["EMP011", "Karen", "Thomas", "Sales", "Account Executive", "2017-03-22", 9, 80000, 20000, 100000, 4.4, "Mark Wilson", "Dallas", 1502, "karen.thomas@company.com", 22, 1, 11, 3, 10, 8, 7200, 600, 1, "President's Club 2025", "Active"],
    ["EMP012", "Leo", "Jackson", "Operations", "Operations Manager", "2014-10-08", 12, 92000, 10000, 102000, 4.1, "Patricia Adams", "Chicago", 1601, "leo.jackson@company.com", 15, 3, 20, 4, 16, 5, 3000, 1100, 3, "", "Active"],
    ["EMP013", "Maria", "White", "Engineering", "Frontend Developer", "2021-06-15", 5, 88000, 9000, 97000, 4.2, "Robert Chen", "New York", 1206, "maria.white@company.com", 42, 4, 11, 2, 6, 18, 1600, 1800, 4, "", "Active"],
    ["EMP014", "Nathan", "Harris", "Finance", "Budget Analyst", "2019-01-07", 7, 76000, 6500, 82500, 3.7, "James Park", "New York", 1107, "nathan.harris@company.com", 25, 3, 8, 7, 9, 10, 1000, 900, 3, "", "Active"],
    ["EMP015", "Olivia", "Clark", "Human Resources", "Recruiter", "2022-09-19", 4, 65000, 5000, 70000, 3.6, "Linda Torres", "San Francisco", 1402, "olivia.clark@company.com", 20, 1, 5, 3, 4, 2, 2800, 500, 1, "", "Active"],
    ["EMP016", "Peter", "Lewis", "Engineering", "Data Scientist", "2020-05-11", 6, 110000, 15000, 125000, 4.7, "Robert Chen", "San Francisco", 1207, "peter.lewis@company.com", 55, 6, 14, 1, 12, 22, 2400, 3000, 5, "", "Active"],
    ["EMP017", "Quinn", "Robinson", "Marketing", "Graphic Designer", "2018-08-25", 8, 70000, 6000, 76000, 3.9, "Sarah Lee", "Chicago", 1304, "quinn.robinson@company.com", 35, 2, 9, 4, 8, 0, 500, 2200, 3, "", "Active"],
    ["EMP018", "Rachel", "Walker", "Sales", "Regional Director", "2013-02-14", 13, 120000, 25000, 145000, 4.9, "Mark Wilson", "Dallas", 1503, "rachel.walker@company.com", 12, 4, 22, 1, 18, 15, 9500, 800, 2, "Exceeded targets 5 consecutive quarters", "Active"],
    ["EMP019", "Samuel", "Young", "Operations", "Logistics Coordinator", "2021-11-30", 5, 62000, 4500, 66500, 3.5, "Patricia Adams", "Chicago", 1602, "samuel.young@company.com", 18, 1, 6, 8, 3, 12, 1200, 400, 1, "", "Active"],
    ["EMP020", "Tina", "King", "Engineering", "Security Engineer", "2018-04-17", 8, 108000, 13000, 121000, 4.5, "Robert Chen", "New York", 1208, "tina.king@company.com", 48, 8, 13, 2, 11, 20, 1900, 2000, 4, "", "Active"],
    ["EMP021", "Ulysses", "Wright", "Finance", "Tax Specialist", "2016-12-05", 10, 90000, 9000, 99000, 4.0, "James Park", "New York", 1108, "ulysses.wright@company.com", 22, 5, 12, 3, 13, 6, 1700, 800, 3, "", "Active"],
    ["EMP022", "Victoria", "Lopez", "Marketing", "SEO Analyst", "2023-01-09", 3, 67000, 5000, 72000, 3.8, "Sarah Lee", "San Francisco", 1305, "victoria.lopez@company.com", 30, 2, 4, 2, 2, 8, 900, 1100, 2, "", "Active"],
    ["EMP023", "William", "Hill", "Sales", "Business Development", "2019-07-21", 7, 85000, 18000, 103000, 4.3, "Mark Wilson", "Dallas", 1504, "william.hill@company.com", 16, 2, 10, 5, 7, 6, 5800, 700, 1, "", "Active"],
    ["EMP024", "Xena", "Scott", "Human Resources", "Training Manager", "2017-05-30", 9, 82000, 8500, 90500, 4.1, "Linda Torres", "Chicago", 1403, "xena.scott@company.com", 60, 4, 15, 2, 14, 0, 3500, 600, 2, "", "Active"],
    ["EMP025", "Yusuf", "Green", "Operations", "Supply Chain Analyst", "2020-10-12", 6, 74000, 6000, 80000, 3.7, "Patricia Adams", "New York", 1603, "yusuf.green@company.com", 26, 3, 7, 4, 6, 14, 2100, 900, 2, "", "Active"],
]

for r_idx, row_data in enumerate(data_rows, 2):
    for c_idx, value in enumerate(row_data, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Set reasonable column widths
ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 16
ws.column_dimensions["E"].width = 22
ws.column_dimensions["F"].width = 12
ws.column_dimensions["O"].width = 30

wb.save(output_path)
print(f"File saved to {output_path}")

# Launch LibreOffice Calc
env = os.environ.copy()
env["DISPLAY"] = ":0"
subprocess.Popen(
    ["libreoffice", "--calc", output_path],
    stdout=subprocess.DEVNULL,
    stderr=subprocess.DEVNULL,
    env=env,
)
time.sleep(2)
print("LibreOffice Calc launched.")
