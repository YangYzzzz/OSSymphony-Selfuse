"""
Initial Setup: Fixed-width text-to-columns task
Task ID: calc_dop_texttocol_fixedwidth_043
Domain: libreoffice_calc

Creates a spreadsheet with a 'LegacyImport' sheet containing fixed-width
records in column A where positions 1-6 are Employee ID and positions 7-20
are Name (padded with spaces). Column B is empty. Columns C-E contain
additional employee data.
"""

import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_dop_texttocol_fixedwidth_043'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: LegacyImport ---
    ws = wb.active
    ws.title = 'LegacyImport'

    # Row 1: fixed-width header (20 chars: 'EMP-ID' in positions 1-6, 'NAME' starting at 7)
    ws['A1'] = 'EMP-IDNAME          '  # 20-char header: positions 1-6='EMP-ID', 7-20='NAME          '

    # Columns C-E: additional data headers
    ws['C1'] = 'Department'
    ws['D1'] = 'HireDate'
    ws['E1'] = 'Status'

    # Style the header row
    for col_letter in ['A', 'C', 'D', 'E']:
        cell = ws[f'{col_letter}1']
        cell.font = Font(bold=True)

    # Fixed-width records: 40 rows of 20-char strings
    # Format: positions 1-6 = Employee ID (6 chars), positions 7-20 = Name (14 chars padded)
    records = [
        'EMP001John Smith     ',   # EMP001 | John Smith
        'EMP002Maria Garcia   ',   # EMP002 | Maria Garcia
        'EMP003James Wilson   ',   # EMP003 | James Wilson
        'EMP004Linda Chen     ',   # EMP004 | Linda Chen
        'EMP005Robert Davis   ',   # EMP005 | Robert Davis
        'EMP006Sarah Thompson ',   # EMP006 | Sarah Thompson
        'EMP007Michael Brown  ',   # EMP007 | Michael Brown
        'EMP008Jessica Lee    ',   # EMP008 | Jessica Lee
        'EMP009David Martinez ',   # EMP009 | David Martinez
        'EMP010Emily Johnson  ',   # EMP010 | Emily Johnson
        'EMP011Daniel White   ',   # EMP011 | Daniel White
        'EMP012Ashley Harris  ',   # EMP012 | Ashley Harris
        'EMP013Matthew Clark  ',   # EMP013 | Matthew Clark
        'EMP014Amanda Lewis   ',   # EMP014 | Amanda Lewis
        'EMP015Christopher R  ',   # EMP015 | Christopher R
        'EMP016Stephanie Hall ',   # EMP016 | Stephanie Hall
        'EMP017Joshua Allen   ',   # EMP017 | Joshua Allen
        'EMP018Megan Young    ',   # EMP018 | Megan Young
        'EMP019Andrew King    ',   # EMP019 | Andrew King
        'EMP020Kimberly Wright',   # EMP020 | Kimberly Wright
        'EMP021Ryan Scott     ',   # EMP021 | Ryan Scott
        'EMP022Brittany Green ',   # EMP022 | Brittany Green
        'EMP023Nicholas Adams ',   # EMP023 | Nicholas Adams
        'EMP024Samantha Baker ',   # EMP024 | Samantha Baker
        'EMP025Tyler Nelson   ',   # EMP025 | Tyler Nelson
        'EMP026Rachel Carter  ',   # EMP026 | Rachel Carter
        'EMP027Brandon Mitchell',  # EMP027 | Brandon Mitchell  (NOTE: exactly 20 chars)
        'EMP028Lauren Perez   ',   # EMP028 | Lauren Perez
        'EMP029Justin Roberts ',   # EMP029 | Justin Roberts
        'EMP030Heather Turner ',   # EMP030 | Heather Turner
        'EMP031Kevin Phillips ',   # EMP031 | Kevin Phillips
        'EMP032Amy Campbell   ',   # EMP032 | Amy Campbell
        'EMP033Brian Parker   ',   # EMP033 | Brian Parker
        'EMP034Angela Evans   ',   # EMP034 | Angela Evans
        'EMP035Nathan Edwards ',   # EMP035 | Nathan Edwards
        'EMP036Amber Collins  ',   # EMP036 | Amber Collins
        'EMP037Aaron Stewart  ',   # EMP037 | Aaron Stewart
        'EMP038Rebecca Sanchez',   # EMP038 | Rebecca Sanchez
        'EMP039Patrick Morris ',   # EMP039 | Patrick Morris
        'EMP040Melissa Rogers ',   # EMP040 | Melissa Rogers
    ]

    # Additional data for columns C-E
    departments = [
        'Engineering', 'Marketing', 'Finance', 'HR', 'Engineering',
        'Sales', 'Engineering', 'Marketing', 'Finance', 'Operations',
        'Engineering', 'HR', 'Sales', 'Finance', 'Engineering',
        'Marketing', 'Operations', 'HR', 'Sales', 'Engineering',
        'Finance', 'Marketing', 'Engineering', 'HR', 'Sales',
        'Finance', 'Operations', 'Marketing', 'Engineering', 'HR',
        'Sales', 'Finance', 'Engineering', 'Marketing', 'Operations',
        'HR', 'Sales', 'Finance', 'Engineering', 'Marketing',
    ]
    hire_dates = [
        '2019-03-15', '2020-06-01', '2018-11-20', '2021-02-10', '2017-09-05',
        '2022-01-15', '2019-07-22', '2020-03-08', '2018-05-14', '2021-08-30',
        '2017-04-17', '2022-05-25', '2019-10-03', '2020-11-18', '2018-02-28',
        '2021-04-12', '2019-06-07', '2020-09-21', '2017-12-04', '2022-03-16',
        '2018-08-09', '2021-01-27', '2019-05-13', '2020-07-06', '2017-10-31',
        '2022-06-20', '2018-03-24', '2021-09-11', '2019-01-29', '2020-04-15',
        '2018-07-03', '2021-11-08', '2017-06-19', '2022-02-24', '2019-08-16',
        '2020-10-02', '2018-01-11', '2021-06-30', '2017-03-22', '2022-07-05',
    ]
    statuses = ['Active', 'Active', 'Active', 'Active', 'Active',
                'Active', 'On Leave', 'Active', 'Active', 'Active',
                'Inactive', 'Active', 'Active', 'Active', 'Active',
                'On Leave', 'Active', 'Active', 'Active', 'Active',
                'Active', 'Active', 'Inactive', 'Active', 'Active',
                'Active', 'Active', 'On Leave', 'Active', 'Active',
                'Active', 'Active', 'Active', 'Inactive', 'Active',
                'Active', 'Active', 'Active', 'Active', 'On Leave']

    # Verify all records are exactly 20 characters, pad/truncate if needed
    for i, rec in enumerate(records):
        # Ensure exactly 20 chars
        rec = rec[:20].ljust(20)
        records[i] = rec

    # Write records to column A, rows 2-41
    # Column B is intentionally left empty (the task will split into B)
    for i, rec in enumerate(records):
        row = i + 2
        ws.cell(row=row, column=1, value=rec)  # Column A: fixed-width string
        # Column B: intentionally empty
        ws.cell(row=row, column=3, value=departments[i])  # Column C
        ws.cell(row=row, column=4, value=hire_dates[i])   # Column D
        ws.cell(row=row, column=5, value=statuses[i])     # Column E

    # Set column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10

    # Save file
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: LegacyImport')
    print(f'  Rows: 1 header + 40 data rows')
    print(f'  Column A: 20-char fixed-width strings (EmpID[6] + Name[14])')
    print(f'  Column B: empty (to be filled by text-to-columns)')
    print(f'  Columns C-E: Department, HireDate, Status')


create_initial()
