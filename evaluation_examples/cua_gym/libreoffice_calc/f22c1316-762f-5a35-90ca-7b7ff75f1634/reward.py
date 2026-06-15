"""
Reward Script: Fixed-width text-to-columns split at position 6
Task ID: calc_dop_texttocol_fixedwidth_043
Domain: libreoffice_calc
Scoring:
  - Component 1: Column A contains only 6-character Employee IDs (not original 20-char strings) (0.4 pts)
  - Component 2: Column B populated with trimmed Name values (was entirely empty in initial) (0.4 pts)
  - Component 3: Spot-check of 6 specific rows match expected EMP IDs and Names exactly (0.2 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_dop_texttocol_fixedwidth_043'
SHEET_NAME = 'LegacyImport'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    The task: column A had 20-char fixed-width strings (positions 1-6=EmpID, 7-20=Name).
    After text-to-columns split at position 6:
    - Column A should contain only 6-character Employee IDs
    - Column B should contain trimmed Name values (previously empty)
    - All 41 rows (1 header + 40 data rows) should be correctly split
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: sheet 'LegacyImport' must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Component 1: Column A cells are 6 characters (split was performed correctly)
    # In the initial file, column A contains 20-character strings (unsplit).
    # In the golden file, column A contains 6-character Employee IDs.
    # This check FAILS on initial (cells are 20 chars long) and PASSES on golden (cells are 6 chars).
    try:
        col_a_long_count = 0  # cells with more than 6 characters (unsplit)
        col_a_six_char_count = 0  # cells with exactly 6 characters (correctly split)
        col_a_nonempty_count = 0  # non-None cells in col A

        for row in range(1, 42):  # rows 1-41 (header + 40 data rows)
            a_val = ws.cell(row=row, column=1).value
            if a_val is not None:
                col_a_nonempty_count += 1
                a_len = len(str(a_val))
                if a_len > 6:
                    col_a_long_count += 1
                elif a_len <= 6:
                    col_a_six_char_count += 1

        if col_a_long_count == 0 and col_a_six_char_count >= 41:
            print(f"PASS: Component 1 — Column A has 6-char Employee IDs in all 41 rows")
            total_score += 0.4
        elif col_a_long_count == 0 and col_a_six_char_count >= 20:
            ratio = col_a_six_char_count / 41
            print(f"PARTIAL: Component 1 — {col_a_six_char_count}/41 rows split correctly")
            total_score += round(ratio * 0.4, 2)  # partial credit proportional to rows split
        else:
            print(f"FAIL: Component 1 — Column A still has {col_a_long_count} unsplit 20-char strings")

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Column B populated with Name values
    # In the initial file, column B is entirely empty (all None).
    # In the golden file, column B has trimmed Name strings in all 41 rows.
    # This check FAILS on initial (B is all None) and PASSES on golden (B has names).
    try:
        col_b_populated_count = 0
        col_b_trimmed_count = 0

        for row in range(1, 42):  # header + 40 data rows
            b_val = ws.cell(row=row, column=2).value
            if b_val is not None:
                col_b_populated_count += 1
                b_str = str(b_val)
                # Trailing spaces trimmed means b_str equals its rstripped version
                if b_str == b_str.rstrip():
                    col_b_trimmed_count += 1

        if col_b_populated_count >= 41:
            print(f"PASS: Component 2 — Column B populated in all 41 rows "
                  f"({col_b_trimmed_count}/41 trimmed correctly)")
            total_score += 0.4
        elif col_b_populated_count >= 20:
            ratio = col_b_populated_count / 41
            print(f"PARTIAL: Component 2 — Column B: {col_b_populated_count}/41 rows populated ({ratio:.0%})")
            total_score += round(ratio * 0.4, 2)
        else:
            print(f"FAIL: Component 2 — Column B has only {col_b_populated_count}/41 rows populated "
                  f"(was entirely empty before task; expected all 41 to be filled)")

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Spot-check 6 specific rows for correct split at position 6
    # Verify header (row 1) and 5 data rows match known expected values.
    # This FAILS on initial (combined strings in A, empty B) and PASSES on golden.
    try:
        spot_checks = [
            # (row, expected_a, expected_b)
            (1, 'EMP-ID', 'NAME'),
            (2, 'EMP001', 'John Smith'),
            (3, 'EMP002', 'Maria Garcia'),
            (7, 'EMP006', 'Sarah Thompson'),
            (10, 'EMP009', 'David Martinez'),
            (11, 'EMP010', 'Emily Johnson'),
        ]

        spot_passed = 0
        spot_total = len(spot_checks)

        for (row, exp_a, exp_b) in spot_checks:
            a_val = ws.cell(row=row, column=1).value
            b_val = ws.cell(row=row, column=2).value

            a_ok = (a_val is not None and str(a_val).strip() == exp_a)
            b_ok = (b_val is not None and str(b_val).strip() == exp_b)

            if a_ok and b_ok:
                spot_passed += 1
            else:
                a_found = repr(str(a_val).strip()) if a_val else 'None'
                b_found = repr(str(b_val).strip()) if b_val else 'None'
                if not a_ok:
                    print(f"FAIL: Component 3 row {row}: A expected '{exp_a}', found {a_found}")
                if not b_ok:
                    print(f"FAIL: Component 3 row {row}: B expected '{exp_b}', found {b_found}")

        if spot_passed == spot_total:
            print(f"PASS: Component 3 — All {spot_total} spot-checks passed (correct EMP IDs and Names)")
            total_score += 0.2
        elif spot_passed > 0:
            ratio = spot_passed / spot_total
            print(f"PARTIAL: Component 3 — {spot_passed}/{spot_total} spot-checks passed")
            total_score += round(ratio * 0.2, 2)
        else:
            print(f"FAIL: Component 3 — 0/{spot_total} spot-checks passed")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
