"""
Initial Setup: Apply diagonal border lines to mark void area in planning grid
Task ID: calc_gg1_022
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_022'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Draft (planning grid) ---
    ws = wb.active
    ws.title = 'Draft'

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18

    # Header row styling
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    # Headers: Row 1
    headers = ['Time Slot', 'Monday', 'Tuesday', 'Wednesday', 'Thursday']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Row 2-11: Time slots from 8:00 AM to 5:00 PM
    time_slots = [
        '8:00 AM', '9:00 AM', '10:00 AM', '11:00 AM', '12:00 PM',
        '1:00 PM', '2:00 PM', '3:00 PM', '4:00 PM', '5:00 PM'
    ]

    # Planning data - B3:E10 is the "cancelled" block (rows 3-10, cols B-E)
    # Row 2 (8:00 AM) and Row 11 (5:00 PM) are outside the void block
    planning_data = {
        # Row 2 (8:00 AM) - outside void block
        (2, 2): 'Team Standup',
        (2, 3): 'Client Call - Apex Inc.',
        (2, 4): 'Sprint Planning',
        (2, 5): 'Budget Review',
        # Row 3 (9:00 AM) - inside void block
        (3, 2): 'Workshop: Q3 Strategy',
        (3, 3): 'Vendor Meeting',
        (3, 4): 'Design Review',
        (3, 5): 'HR Orientation',
        # Row 4 (10:00 AM)
        (4, 2): 'Product Demo',
        (4, 3): 'Code Review Session',
        (4, 4): 'Marketing Sync',
        (4, 5): 'Finance Check-in',
        # Row 5 (11:00 AM)
        (5, 2): 'Architecture Discussion',
        (5, 3): 'Lunch & Learn',
        (5, 4): '1:1 with Sarah Chen',
        (5, 5): 'Stakeholder Update',
        # Row 6 (12:00 PM)
        (6, 2): 'Lunch Break',
        (6, 3): 'Lunch Break',
        (6, 4): 'Lunch Break',
        (6, 5): 'Lunch Break',
        # Row 7 (1:00 PM)
        (7, 2): 'UX Research Debrief',
        (7, 3): 'Platform Migration',
        (7, 4): 'Security Audit Prep',
        (7, 5): 'Data Pipeline Review',
        # Row 8 (2:00 PM)
        (8, 2): 'Customer Feedback',
        (8, 3): 'Release Planning',
        (8, 4): 'QA Regression Test',
        (8, 5): 'Infra Cost Analysis',
        # Row 9 (3:00 PM)
        (9, 2): 'API Design Workshop',
        (9, 3): 'Team Retro',
        (9, 4): 'Partner Integration',
        (9, 5): 'Compliance Training',
        # Row 10 (4:00 PM)
        (10, 2): 'Sprint Review',
        (10, 3): 'Ops Handoff',
        (10, 4): 'Feature Triage',
        (10, 5): 'Exec Summary Prep',
        # Row 11 (5:00 PM) - outside void block
        (11, 2): 'End-of-Day Wrap-up',
        (11, 3): 'Status Report Filing',
        (11, 4): 'Backlog Grooming',
        (11, 5): 'Weekly Digest',
    }

    # Fill time slot column and planning data
    time_font = Font(name='Arial', size=10, bold=True)
    data_font = Font(name='Arial', size=10)
    data_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for r, ts in enumerate(time_slots, 2):
        cell = ws.cell(row=r, column=1, value=ts)
        cell.font = time_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for (r, c), val in planning_data.items():
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = data_font
        cell.alignment = data_align

    # Add thin borders to all data cells for the grid look
    thin = Side(style='thin', color='000000')
    grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(1, 12):
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = grid_border

    # Row heights
    ws.row_dimensions[1].height = 25
    for r in range(2, 12):
        ws.row_dimensions[r].height = 30

    # Additional column F with notes (outside void area)
    ws.cell(row=1, column=6, value='Notes').font = header_font
    ws.cell(row=1, column=6).fill = header_fill
    ws.cell(row=1, column=6).alignment = header_align
    ws.cell(row=2, column=6, value='High priority week')
    ws.cell(row=3, column=6, value='Cancelled - office closure')
    ws.cell(row=11, column=6, value='Normal schedule resumes')

    # --- Sheet 2: Notes ---
    ws2 = wb.create_sheet('Notes')
    ws2['A1'] = 'Planning Notes'
    ws2['A1'].font = Font(name='Arial', size=14, bold=True)
    ws2['A3'] = 'Week of March 17-21, 2025'
    ws2['A4'] = 'The block B3:E10 covers cancelled time slots due to office renovation.'
    ws2['A5'] = 'These slots should be marked as void in the Draft sheet.'
    ws2['A7'] = 'Contact: Marcus Rivera, Facilities Manager'
    ws2['A8'] = 'Expected completion: March 24, 2025'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
