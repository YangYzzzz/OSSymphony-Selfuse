"""
Reward Script: Apply diagonal border lines (both diagonals) to B3:E10 on 'Draft' sheet
Task ID: calc_gg1_022
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): All cells in B3:E10 have diagonalDown=True with a diagonal style
  Component 2 (0.4): All cells in B3:E10 have diagonalUp=True with a diagonal style
  Component 3 (0.2): Cells outside B3:E10 are unaffected (no diagonal borders added)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_022'


def persist_app_state(domain):
    """Try to save any unsaved LibreOffice state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify that diagonal border lines (both diagonals) are applied to B3:E10 on 'Draft' sheet.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Draft' sheet must exist
    if 'Draft' not in wb.sheetnames:
        print(f"CRITICAL: 'Draft' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Draft']

    # Target range: B3:E10 (rows 3-10, columns 2-5) = 32 cells total
    TARGET_MIN_ROW = 3
    TARGET_MAX_ROW = 10
    TARGET_MIN_COL = 2  # B
    TARGET_MAX_COL = 5  # E
    TOTAL_TARGET_CELLS = (TARGET_MAX_ROW - TARGET_MIN_ROW + 1) * (TARGET_MAX_COL - TARGET_MIN_COL + 1)

    # Component 1: diagonalDown=True with diagonal style on all cells in B3:E10 (0.4 points)
    # This checks the top-left to bottom-right diagonal line
    try:
        down_pass_count = 0
        for row in range(TARGET_MIN_ROW, TARGET_MAX_ROW + 1):
            for col in range(TARGET_MIN_COL, TARGET_MAX_COL + 1):
                cell = ws.cell(row=row, column=col)
                b = cell.border
                has_diagonal_style = (b.diagonal is not None and
                                      b.diagonal.style is not None)
                has_diag_down = getattr(b, 'diagonalDown', False) is True
                if has_diagonal_style and has_diag_down:
                    down_pass_count += 1

        ratio_down = down_pass_count / TOTAL_TARGET_CELLS
        if ratio_down == 1.0:
            print(f"PASS: Component 1 — All {TOTAL_TARGET_CELLS} cells have diagonalDown border (0.4 pts)")
            total_score += 0.4
        elif ratio_down > 0:
            partial = round(0.4 * ratio_down, 2)
            print(f"PARTIAL: Component 1 — {down_pass_count}/{TOTAL_TARGET_CELLS} cells have diagonalDown ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No cells in B3:E10 have diagonalDown border")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: diagonalUp=True with diagonal style on all cells in B3:E10 (0.4 points)
    # This checks the top-right to bottom-left diagonal line
    try:
        up_pass_count = 0
        for row in range(TARGET_MIN_ROW, TARGET_MAX_ROW + 1):
            for col in range(TARGET_MIN_COL, TARGET_MAX_COL + 1):
                cell = ws.cell(row=row, column=col)
                b = cell.border
                has_diagonal_style = (b.diagonal is not None and
                                      b.diagonal.style is not None)
                has_diag_up = getattr(b, 'diagonalUp', False) is True
                if has_diagonal_style and has_diag_up:
                    up_pass_count += 1

        ratio_up = up_pass_count / TOTAL_TARGET_CELLS
        if ratio_up == 1.0:
            print(f"PASS: Component 2 — All {TOTAL_TARGET_CELLS} cells have diagonalUp border (0.4 pts)")
            total_score += 0.4
        elif ratio_up > 0:
            partial = round(0.4 * ratio_up, 2)
            print(f"PARTIAL: Component 2 — {up_pass_count}/{TOTAL_TARGET_CELLS} cells have diagonalUp ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No cells in B3:E10 have diagonalUp border")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Diagonals applied ONLY to B3:E10 — outside cells unaffected (0.2 points)
    # This is a compound check: B3:E10 must HAVE diagonals AND outside cells must NOT.
    # The B3:E10 requirement anchors this to the task change, preventing score on initial_env.
    try:
        # First: at least 1 cell in B3:E10 must have diagonal borders (gate)
        any_inside_has_diagonal = (down_pass_count > 0 or up_pass_count > 0)

        if not any_inside_has_diagonal:
            print(f"FAIL: Component 3 — No diagonal borders in B3:E10, cannot verify selectivity")
        else:
            outside_cells = [
                (1, 1),   # A1
                (2, 2),   # B2
                (2, 5),   # E2
                (3, 1),   # A3
                (3, 6),   # F3
                (10, 1),  # A10
                (10, 6),  # F10
                (11, 2),  # B11
                (11, 5),  # E11
                (1, 3),   # C1
            ]
            outside_violations = 0
            for r, c in outside_cells:
                cell = ws.cell(row=r, column=c)
                b = cell.border
                has_diagonal_style = (b.diagonal is not None and
                                      b.diagonal.style is not None)
                has_diag_down = getattr(b, 'diagonalDown', False) is True
                has_diag_up = getattr(b, 'diagonalUp', False) is True
                if has_diagonal_style and (has_diag_down or has_diag_up):
                    outside_violations += 1
                    print(f"  VIOLATION: Cell ({r},{c}) has diagonal border but is outside B3:E10")

            if outside_violations == 0:
                print(f"PASS: Component 3 — Diagonals in B3:E10 and none outside (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 3 — {outside_violations} cells outside B3:E10 have diagonal borders")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist app state before verification (best-effort)
persist_app_state("libreoffice_calc")

# Run verification
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
