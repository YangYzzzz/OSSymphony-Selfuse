"""
Initial Setup: Financial model with all cells locked (Protected), including 5 yellow input cells
Task ID: calc_cop_protection_003
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection

WORKDIR = '/home/user'
TASK_ID = 'calc_cop_protection_003'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

# Colors
YELLOW_FILL = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
HEADER_FILL = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
LABEL_FILL = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
FORMULA_FILL = PatternFill(start_color='FFE2EFDA', end_color='FFE2EFDA', fill_type='solid')

HEADER_FONT = Font(name='Calibri', size=13, bold=True, color='FFFFFFFF')
LABEL_FONT = Font(name='Calibri', size=11, bold=True, color='FF1F3864')
FORMULA_FONT = Font(name='Calibri', size=11, italic=True, color='FF375623')
INPUT_FONT = Font(name='Calibri', size=11, color='FF7B3F00')
NORMAL_FONT = Font(name='Calibri', size=11)

LOCKED = Protection(locked=True, hidden=False)


def thin_border():
    thin = Side(style='thin', color='FF9DC3E6')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'FinancialModel'

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18

    # --- Row 1: Title ---
    ws.merge_cells('A1:C1')
    ws['A1'] = 'Annual Financial Projection Model'
    ws['A1'].font = Font(name='Calibri', size=15, bold=True, color='FFFFFFFF')
    ws['A1'].fill = PatternFill(start_color='FF1F3864', end_color='FF1F3864', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].protection = LOCKED
    ws.row_dimensions[1].height = 32

    # --- Row 2: Subtitle ---
    ws.merge_cells('A2:C2')
    ws['A2'] = 'FY 2025 — Confidential Internal Use Only'
    ws['A2'].font = Font(name='Calibri', size=11, italic=True, color='FF7F7F7F')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].protection = LOCKED
    ws.row_dimensions[2].height = 20

    # --- Row 3: blank separator ---
    ws.row_dimensions[3].height = 8

    # --- Section 1: Revenue Inputs (rows 4-7) ---
    # Row 4: section header
    ws.merge_cells('A4:C4')
    ws['A4'] = 'SECTION 1: Revenue Inputs'
    ws['A4'].font = HEADER_FONT
    ws['A4'].fill = HEADER_FILL
    ws['A4'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws['A4'].protection = LOCKED
    ws.row_dimensions[4].height = 24

    # Row 5: input cell (yellow) — B5 is a USER INPUT
    ws['A5'] = 'Projected Annual Revenue ($)'
    ws['A5'].font = LABEL_FONT
    ws['A5'].fill = LABEL_FILL
    ws['A5'].alignment = Alignment(vertical='center', indent=1)
    ws['A5'].protection = LOCKED
    ws['B5'] = 4250000
    ws['B5'].font = INPUT_FONT
    ws['B5'].fill = YELLOW_FILL
    ws['B5'].number_format = '$#,##0'
    ws['B5'].alignment = Alignment(horizontal='right', vertical='center')
    ws['B5'].protection = LOCKED   # Initially locked — task is to unlock this
    ws['C5'] = 'Enter projected revenue'
    ws['C5'].font = Font(name='Calibri', size=10, italic=True, color='FF767171')
    ws['C5'].protection = LOCKED
    ws.row_dimensions[5].height = 22

    # Row 6: formula cell
    ws['A6'] = 'Monthly Revenue Run Rate ($)'
    ws['A6'].font = LABEL_FONT
    ws['A6'].fill = LABEL_FILL
    ws['A6'].alignment = Alignment(vertical='center', indent=1)
    ws['A6'].protection = LOCKED
    ws['B6'] = '=B5/12'
    ws['B6'].font = FORMULA_FONT
    ws['B6'].fill = FORMULA_FILL
    ws['B6'].number_format = '$#,##0'
    ws['B6'].alignment = Alignment(horizontal='right', vertical='center')
    ws['B6'].protection = LOCKED
    ws['C6'] = 'Auto-calculated'
    ws['C6'].font = Font(name='Calibri', size=10, italic=True, color='FF767171')
    ws['C6'].protection = LOCKED
    ws.row_dimensions[6].height = 22

    # Row 7: blank separator
    ws.row_dimensions[7].height = 8

    # --- Section 2: Cost Structure (rows 8-10) ---
    # Row 8: section header + input
    ws['A8'] = 'SECTION 2: Cost Structure'
    ws['A8'].font = HEADER_FONT
    ws['A8'].fill = HEADER_FILL
    ws['A8'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws['A8'].protection = LOCKED
    ws.merge_cells('A8:C8')
    ws.row_dimensions[8].height = 24

    # Row 9: input cell (yellow) — B9 is a USER INPUT (wait, task says B8...)
    # Per task: B5, B8, B11, B14, B17 are the input cells
    # Let's restructure: section headers are rows 4, 7, 10, 13, 16 and inputs are at rows 5, 8, 11, 14, 17
    # Re-plan: I'll unmerge and redo to match the exact row spec

    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = 'FinancialModel'

    # Column widths
    ws2.column_dimensions['A'].width = 32
    ws2.column_dimensions['B'].width = 22
    ws2.column_dimensions['C'].width = 20

    def set_cell(cell, value=None, font=None, fill=None, alignment=None, number_format=None, protection=None, border=None):
        if value is not None:
            cell.value = value
        if font is not None:
            cell.font = font
        if fill is not None:
            cell.fill = fill
        if alignment is not None:
            cell.alignment = alignment
        if number_format is not None:
            cell.number_format = number_format
        if protection is not None:
            cell.protection = protection
        if border is not None:
            cell.border = border

    # Row 1: Title
    ws2.merge_cells('A1:C1')
    set_cell(ws2['A1'],
             value='Annual Financial Projection Model — FY 2025',
             font=Font(name='Calibri', size=15, bold=True, color='FFFFFFFF'),
             fill=PatternFill(start_color='FF1F3864', end_color='FF1F3864', fill_type='solid'),
             alignment=Alignment(horizontal='center', vertical='center'),
             protection=LOCKED)
    ws2.row_dimensions[1].height = 34

    # Row 2: Instructions
    ws2.merge_cells('A2:C2')
    set_cell(ws2['A2'],
             value='Yellow cells are user inputs. All formulas are auto-calculated. Do not modify formula cells.',
             font=Font(name='Calibri', size=10, italic=True, color='FF595959'),
             alignment=Alignment(horizontal='center', vertical='center'),
             protection=LOCKED)
    ws2.row_dimensions[2].height = 18

    # Row 3: blank
    ws2.row_dimensions[3].height = 10

    # Row 4: Section 1 header
    ws2.merge_cells('A4:C4')
    set_cell(ws2['A4'],
             value='SECTION 1 — Revenue Assumptions',
             font=Font(name='Calibri', size=12, bold=True, color='FFFFFFFF'),
             fill=PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid'),
             alignment=Alignment(horizontal='left', vertical='center', indent=1),
             protection=LOCKED)
    ws2.row_dimensions[4].height = 26

    # Row 5: INPUT CELL — B5 (yellow)
    set_cell(ws2['A5'],
             value='Projected Annual Gross Revenue ($)',
             font=Font(name='Calibri', size=11, bold=True, color='FF1F3864'),
             fill=PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid'),
             alignment=Alignment(vertical='center', indent=1),
             protection=LOCKED)
    set_cell(ws2['B5'],
             value=4800000,
             font=Font(name='Calibri', size=11, bold=True, color='FF7B3F00'),
             fill=PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid'),
             number_format='$#,##0',
             alignment=Alignment(horizontal='right', vertical='center'),
             protection=LOCKED)
    set_cell(ws2['C5'],
             value='Input: Annual Revenue',
             font=Font(name='Calibri', size=10, italic=True, color='FF767171'),
             protection=LOCKED)
    ws2.row_dimensions[5].height = 22

    # Row 6: Formula — monthly revenue
    set_cell(ws2['A6'],
             value='Monthly Revenue Run Rate ($)',
             font=Font(name='Calibri', size=11, color='FF1F3864'),
             fill=PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid'),
             alignment=Alignment(vertical='center', indent=1),
             protection=LOCKED)
    set_cell(ws2['B6'],
             value='=B5/12',
             font=Font(name='Calibri', size=11, italic=True, color='FF375623'),
             fill=PatternFill(start_color='FFE2EFDA', end_color='FFE2EFDA', fill_type='solid'),
             number_format='$#,##0',
             alignment=Alignment(horizontal='right', vertical='center'),
             protection=LOCKED)
    set_cell(ws2['C6'],
             value='= Annual Revenue / 12',
             font=Font(name='Calibri', size=10, italic=True, color='FF767171'),
             protection=LOCKED)
    ws2.row_dimensions[6].height = 22

    # Row 7: blank
    ws2.row_dimensions[7].height = 10

    # Row 8: Section 2 header
    ws2.merge_cells('A8:C8')
    set_cell(ws2['A8'],
             value='SECTION 2 — Operating Cost Parameters',
             font=Font(name='Calibri', size=12, bold=True, color='FFFFFFFF'),
             fill=PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid'),
             alignment=Alignment(horizontal='left', vertical='center', indent=1),
             protection=LOCKED)
    ws2.row_dimensions[8].height = 26

    # Row 9: Sub-label for section 2
    set_cell(ws2['A9'],
             value='Cost of Goods Sold — Rate (%)',
             font=Font(name='Calibri', size=11, color='FF1F3864'),
             fill=PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid'),
             alignment=Alignment(vertical='center', indent=1),
             protection=LOCKED)
    set_cell(ws2['B9'],
             value='=B11*B5',
             font=Font(name='Calibri', size=11, italic=True, color='FF375623'),
             fill=PatternFill(start_color='FFE2EFDA', end_color='FFE2EFDA', fill_type='solid'),
             number_format='$#,##0',
             alignment=Alignment(horizontal='right', vertical='center'),
             protection=LOCKED)
    set_cell(ws2['C9'],
             value='= COGS Rate x Revenue',
             font=Font(name='Calibri', size=10, italic=True, color='FF767171'),
             protection=LOCKED)
    ws2.row_dimensions[9].height = 22

    # Row 10: blank
    ws2.row_dimensions[10].height = 10

    # Row 11: INPUT CELL — B11 (yellow)
    ws2.merge_cells('A10:C10')
    set_cell(ws2['A10'],
             value='COGS & Gross Margin Inputs',
             font=Font(name='Calibri', size=11, bold=True, color='FF595959'),
             alignment=Alignment(horizontal='left', vertical='center', indent=1),
             protection=LOCKED)
    ws2.row_dimensions[10].height = 18

    set_cell(ws2['A11'],
             value='Cost of Goods Sold Rate (% of Revenue)',
             font=Font(name='Calibri', size=11, bold=True, color='FF1F3864'),
             fill=PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid'),
             alignment=Alignment(vertical='center', indent=1),
             protection=LOCKED)
    set_cell(ws2['B11'],
             value=0.42,
             font=Font(name='Calibri', size=11, bold=True, color='FF7B3F00'),
             fill=PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid'),
             number_format='0.00%',
             alignment=Alignment(horizontal='right', vertical='center'),
             protection=LOCKED)
    set_cell(ws2['C11'],
             value='Input: COGS Rate',
             font=Font(name='Calibri', size=10, italic=True, color='FF767171'),
             protection=LOCKED)
    ws2.row_dimensions[11].height = 22

    # Row 12: Formula — gross margin
    set_cell(ws2['A12'],
             value='Gross Profit ($)',
             font=Font(name='Calibri', size=11, color='FF1F3864'),
             fill=PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid'),
             alignment=Alignment(vertical='center', indent=1),
             protection=LOCKED)
    set_cell(ws2['B12'],
             value='=B5*(1-B11)',
             font=Font(name='Calibri', size=11, italic=True, color='FF375623'),
             fill=PatternFill(start_color='FFE2EFDA', end_color='FFE2EFDA', fill_type='solid'),
             number_format='$#,##0',
             alignment=Alignment(horizontal='right', vertical='center'),
             protection=LOCKED)
    set_cell(ws2['C12'],
             value='= Revenue x (1 - COGS Rate)',
             font=Font(name='Calibri', size=10, italic=True, color='FF767171'),
             protection=LOCKED)
    ws2.row_dimensions[12].height = 22

    # Row 13: blank
    ws2.row_dimensions[13].height = 10

    # Row 14: Section 3 header
    ws2.merge_cells('A13:C13')
    set_cell(ws2['A13'],
             value='SECTION 3 — Operating Expenses',
             font=Font(name='Calibri', size=12, bold=True, color='FFFFFFFF'),
             fill=PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid'),
             alignment=Alignment(horizontal='left', vertical='center', indent=1),
             protection=LOCKED)
    ws2.row_dimensions[13].height = 26

    # Row 14: INPUT CELL — B14 (yellow)
    set_cell(ws2['A14'],
             value='Total Operating Expenses ($)',
             font=Font(name='Calibri', size=11, bold=True, color='FF1F3864'),
             fill=PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid'),
             alignment=Alignment(vertical='center', indent=1),
             protection=LOCKED)
    set_cell(ws2['B14'],
             value=1350000,
             font=Font(name='Calibri', size=11, bold=True, color='FF7B3F00'),
             fill=PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid'),
             number_format='$#,##0',
             alignment=Alignment(horizontal='right', vertical='center'),
             protection=LOCKED)
    set_cell(ws2['C14'],
             value='Input: Total OpEx',
             font=Font(name='Calibri', size=10, italic=True, color='FF767171'),
             protection=LOCKED)
    ws2.row_dimensions[14].height = 22

    # Row 15: Formula — operating income
    set_cell(ws2['A15'],
             value='Operating Income (EBIT) ($)',
             font=Font(name='Calibri', size=11, color='FF1F3864'),
             fill=PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid'),
             alignment=Alignment(vertical='center', indent=1),
             protection=LOCKED)
    set_cell(ws2['B15'],
             value='=B12-B14',
             font=Font(name='Calibri', size=11, italic=True, color='FF375623'),
             fill=PatternFill(start_color='FFE2EFDA', end_color='FFE2EFDA', fill_type='solid'),
             number_format='$#,##0',
             alignment=Alignment(horizontal='right', vertical='center'),
             protection=LOCKED)
    set_cell(ws2['C15'],
             value='= Gross Profit - Operating Expenses',
             font=Font(name='Calibri', size=10, italic=True, color='FF767171'),
             protection=LOCKED)
    ws2.row_dimensions[15].height = 22

    # Row 16: blank
    ws2.row_dimensions[16].height = 10

    # Row 17: Section 4 + INPUT CELL — B17 (yellow)
    ws2.merge_cells('A16:C16')
    set_cell(ws2['A16'],
             value='SECTION 4 — Tax & Net Income',
             font=Font(name='Calibri', size=12, bold=True, color='FFFFFFFF'),
             fill=PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid'),
             alignment=Alignment(horizontal='left', vertical='center', indent=1),
             protection=LOCKED)
    ws2.row_dimensions[16].height = 26

    set_cell(ws2['A17'],
             value='Effective Tax Rate (%)',
             font=Font(name='Calibri', size=11, bold=True, color='FF1F3864'),
             fill=PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid'),
             alignment=Alignment(vertical='center', indent=1),
             protection=LOCKED)
    set_cell(ws2['B17'],
             value=0.21,
             font=Font(name='Calibri', size=11, bold=True, color='FF7B3F00'),
             fill=PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid'),
             number_format='0.00%',
             alignment=Alignment(horizontal='right', vertical='center'),
             protection=LOCKED)
    set_cell(ws2['C17'],
             value='Input: Tax Rate',
             font=Font(name='Calibri', size=10, italic=True, color='FF767171'),
             protection=LOCKED)
    ws2.row_dimensions[17].height = 22

    # Row 18: Formula — tax amount
    set_cell(ws2['A18'],
             value='Income Tax Expense ($)',
             font=Font(name='Calibri', size=11, color='FF1F3864'),
             fill=PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid'),
             alignment=Alignment(vertical='center', indent=1),
             protection=LOCKED)
    set_cell(ws2['B18'],
             value='=MAX(0,B15*B17)',
             font=Font(name='Calibri', size=11, italic=True, color='FF375623'),
             fill=PatternFill(start_color='FFE2EFDA', end_color='FFE2EFDA', fill_type='solid'),
             number_format='$#,##0',
             alignment=Alignment(horizontal='right', vertical='center'),
             protection=LOCKED)
    set_cell(ws2['C18'],
             value='= EBIT x Tax Rate (if positive)',
             font=Font(name='Calibri', size=10, italic=True, color='FF767171'),
             protection=LOCKED)
    ws2.row_dimensions[18].height = 22

    # Row 19: Formula — net income
    set_cell(ws2['A19'],
             value='Net Income After Tax ($)',
             font=Font(name='Calibri', size=11, bold=True, color='FF1F3864'),
             fill=PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid'),
             alignment=Alignment(vertical='center', indent=1),
             protection=LOCKED)
    set_cell(ws2['B19'],
             value='=B15-B18',
             font=Font(name='Calibri', size=11, bold=True, italic=True, color='FF375623'),
             fill=PatternFill(start_color='FFE2EFDA', end_color='FFE2EFDA', fill_type='solid'),
             number_format='$#,##0',
             alignment=Alignment(horizontal='right', vertical='center'),
             protection=LOCKED)
    set_cell(ws2['C19'],
             value='= EBIT - Tax Expense',
             font=Font(name='Calibri', size=10, italic=True, color='FF767171'),
             protection=LOCKED)
    ws2.row_dimensions[19].height = 22

    # Row 20: blank
    ws2.row_dimensions[20].height = 10

    # Row 21: Section 5 header
    ws2.merge_cells('A21:C21')
    set_cell(ws2['A21'],
             value='SECTION 5 — Capital Expenditures & Cash Flow',
             font=Font(name='Calibri', size=12, bold=True, color='FFFFFFFF'),
             fill=PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid'),
             alignment=Alignment(horizontal='left', vertical='center', indent=1),
             protection=LOCKED)
    ws2.row_dimensions[21].height = 26

    # Row 22-24: additional data rows
    rows_22_24 = [
        ('Depreciation & Amortization ($)', '=B14*0.08', '= 8% of Operating Expenses'),
        ('Capital Expenditures Budget ($)', 520000, 'Approved FY2025 CapEx'),
        ('Free Cash Flow Estimate ($)', '=B19+B22-B23', '= Net Income + D&A - CapEx'),
    ]
    for i, (label, val, note) in enumerate(rows_22_24, 22):
        set_cell(ws2[f'A{i}'],
                 value=label,
                 font=Font(name='Calibri', size=11, color='FF1F3864'),
                 fill=PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid'),
                 alignment=Alignment(vertical='center', indent=1),
                 protection=LOCKED)
        if isinstance(val, str) and val.startswith('='):
            set_cell(ws2[f'B{i}'],
                     value=val,
                     font=Font(name='Calibri', size=11, italic=True, color='FF375623'),
                     fill=PatternFill(start_color='FFE2EFDA', end_color='FFE2EFDA', fill_type='solid'),
                     number_format='$#,##0',
                     alignment=Alignment(horizontal='right', vertical='center'),
                     protection=LOCKED)
        else:
            set_cell(ws2[f'B{i}'],
                     value=val,
                     font=Font(name='Calibri', size=11, color='FF1F3864'),
                     fill=PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid'),
                     number_format='$#,##0',
                     alignment=Alignment(horizontal='right', vertical='center'),
                     protection=LOCKED)
        set_cell(ws2[f'C{i}'],
                 value=note,
                 font=Font(name='Calibri', size=10, italic=True, color='FF767171'),
                 protection=LOCKED)
        ws2.row_dimensions[i].height = 22

    # Row 25: blank
    ws2.row_dimensions[25].height = 10

    # Row 26: Section 6 header — Summary
    ws2.merge_cells('A26:C26')
    set_cell(ws2['A26'],
             value='SECTION 6 — Key Financial Ratios',
             font=Font(name='Calibri', size=12, bold=True, color='FFFFFFFF'),
             fill=PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid'),
             alignment=Alignment(horizontal='left', vertical='center', indent=1),
             protection=LOCKED)
    ws2.row_dimensions[26].height = 26

    # Row 27-30: financial ratio formulas
    rows_27_30 = [
        ('Gross Margin (%)', '=1-B11', '= 1 - COGS Rate'),
        ('EBIT Margin (%)', '=IF(B5>0,B15/B5,0)', '= EBIT / Revenue'),
        ('Net Profit Margin (%)', '=IF(B5>0,B19/B5,0)', '= Net Income / Revenue'),
        ('Return on Revenue (%)', '=IF(B5>0,(B19+B22)/B5,0)', '= (Net Income + D&A) / Revenue'),
    ]
    for i, (label, val, note) in enumerate(rows_27_30, 27):
        set_cell(ws2[f'A{i}'],
                 value=label,
                 font=Font(name='Calibri', size=11, color='FF1F3864'),
                 fill=PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid'),
                 alignment=Alignment(vertical='center', indent=1),
                 protection=LOCKED)
        set_cell(ws2[f'B{i}'],
                 value=val,
                 font=Font(name='Calibri', size=11, italic=True, color='FF375623'),
                 fill=PatternFill(start_color='FFE2EFDA', end_color='FFE2EFDA', fill_type='solid'),
                 number_format='0.00%',
                 alignment=Alignment(horizontal='right', vertical='center'),
                 protection=LOCKED)
        set_cell(ws2[f'C{i}'],
                 value=note,
                 font=Font(name='Calibri', size=10, italic=True, color='FF767171'),
                 protection=LOCKED)
        ws2.row_dimensions[i].height = 22

    # Note: Total rows = 30 (rows 1-30). Rows 3,7,10,13,16,20,25 are blanks/spacers.
    # Input rows: 5 (B5), 11 (B11), 14 (B14), 17 (B17) — need B8 too!
    # B8 is at row 8 which was set as section header merge. Fix: unmerge row 8,
    # move section header to row 7 (currently blank), and put input at row 8.

    # The sheet currently has section header at row 8 (merged A8:C8).
    # But task requires B8 as an input cell. We need to fix the layout.
    # Let's unmerge A8:C8 and put an input cell at B8 instead.
    ws2.unmerge_cells('A8:C8')

    # Row 7: Make it the section header
    ws2.merge_cells('A7:C7')
    set_cell(ws2['A7'],
             value='SECTION 2 — Operating Cost Parameters',
             font=Font(name='Calibri', size=12, bold=True, color='FFFFFFFF'),
             fill=PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid'),
             alignment=Alignment(horizontal='left', vertical='center', indent=1),
             protection=LOCKED)
    ws2.row_dimensions[7].height = 26

    # Row 8: INPUT CELL — B8 (yellow)
    set_cell(ws2['A8'],
             value='Operating Expense Rate (% of Revenue)',
             font=Font(name='Calibri', size=11, bold=True, color='FF1F3864'),
             fill=PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid'),
             alignment=Alignment(vertical='center', indent=1),
             protection=LOCKED)
    set_cell(ws2['B8'],
             value=0.28,
             font=Font(name='Calibri', size=11, bold=True, color='FF7B3F00'),
             fill=PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid'),
             number_format='0.00%',
             alignment=Alignment(horizontal='right', vertical='center'),
             protection=LOCKED)
    set_cell(ws2['C8'],
             value='Input: OpEx Rate',
             font=Font(name='Calibri', size=10, italic=True, color='FF767171'),
             protection=LOCKED)
    ws2.row_dimensions[8].height = 22

    # Also clear the old merged C8 cell that was part of the section header
    # (unmerge already took care of that)
    # Clear B8 and C8 section header remnants
    ws2['B8'].value = 0.28  # already set above

    # Clear A9 (was 'Cost of Goods Sold — Rate (%)' and B9 was formula, now fix section 2)
    # Actually let's reorganize section 2 properly:
    # Row 7: Section 2 header (done above)
    # Row 8: INPUT B8 (done above)
    # Row 9: Formula using B8 and B5
    set_cell(ws2['A9'],
             value='Total Operating Expenses ($)',
             font=Font(name='Calibri', size=11, color='FF1F3864'),
             fill=PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid'),
             alignment=Alignment(vertical='center', indent=1),
             protection=LOCKED)
    set_cell(ws2['B9'],
             value='=B8*B5',
             font=Font(name='Calibri', size=11, italic=True, color='FF375623'),
             fill=PatternFill(start_color='FFE2EFDA', end_color='FFE2EFDA', fill_type='solid'),
             number_format='$#,##0',
             alignment=Alignment(horizontal='right', vertical='center'),
             protection=LOCKED)
    set_cell(ws2['C9'],
             value='= OpEx Rate x Revenue',
             font=Font(name='Calibri', size=10, italic=True, color='FF767171'),
             protection=LOCKED)

    # Fix section 3 header (was at A13, need to keep correct layout)
    # Row 10: keep as sub-label/blank separator for section 3
    ws2.unmerge_cells('A10:C10')  # was "COGS & Gross Margin Inputs" merged
    ws2.merge_cells('A10:C10')
    set_cell(ws2['A10'],
             value='SECTION 3 — COGS & Gross Margin Inputs',
             font=Font(name='Calibri', size=12, bold=True, color='FFFFFFFF'),
             fill=PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid'),
             alignment=Alignment(horizontal='left', vertical='center', indent=1),
             protection=LOCKED)
    ws2.row_dimensions[10].height = 26

    # Row 13: was section 3 header, now make it section 4 header
    ws2.unmerge_cells('A13:C13')
    ws2.merge_cells('A13:C13')
    set_cell(ws2['A13'],
             value='SECTION 4 — Operating Expenses',
             font=Font(name='Calibri', size=12, bold=True, color='FFFFFFFF'),
             fill=PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid'),
             alignment=Alignment(horizontal='left', vertical='center', indent=1),
             protection=LOCKED)
    ws2.row_dimensions[13].height = 26

    # Row 16: was section 4 header "Tax & Net Income"
    ws2.unmerge_cells('A16:C16')
    ws2.merge_cells('A16:C16')
    set_cell(ws2['A16'],
             value='SECTION 5 — Tax & Net Income',
             font=Font(name='Calibri', size=12, bold=True, color='FFFFFFFF'),
             fill=PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid'),
             alignment=Alignment(horizontal='left', vertical='center', indent=1),
             protection=LOCKED)
    ws2.row_dimensions[16].height = 26

    # Row 21: was section 5 header
    ws2.unmerge_cells('A21:C21')
    ws2.merge_cells('A21:C21')
    set_cell(ws2['A21'],
             value='SECTION 6 — Capital Expenditures & Cash Flow',
             font=Font(name='Calibri', size=12, bold=True, color='FFFFFFFF'),
             fill=PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid'),
             alignment=Alignment(horizontal='left', vertical='center', indent=1),
             protection=LOCKED)

    # Row 26: was section 6 header
    ws2.unmerge_cells('A26:C26')
    ws2.merge_cells('A26:C26')
    set_cell(ws2['A26'],
             value='SECTION 7 — Key Financial Ratios',
             font=Font(name='Calibri', size=12, bold=True, color='FFFFFFFF'),
             fill=PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid'),
             alignment=Alignment(horizontal='left', vertical='center', indent=1),
             protection=LOCKED)

    wb2.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Input cells (yellow, locked): B5, B8, B11, B14, B17')
    print(f'All cells have Protection(locked=True)')


create_initial()
