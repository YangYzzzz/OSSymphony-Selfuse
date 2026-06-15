"""
Reward Script: Unlock only the five yellow-highlighted input cells (B5, B8, B11, B14, B17)
Task ID: calc_cop_protection_003
Domain: libreoffice_calc
Scoring:
  - Component 1: Each of the 5 target cells (B5, B8, B11, B14, B17) is unlocked (locked=False)
                 0.12 pts per cell = 0.60 total for all 5 cells unlocked
  - Component 2: All 5 target cells unlocked AND no non-target cells accidentally unlocked (0.3 pts)
                 This compound check fails on initial (target cells still locked) and passes on golden
  - Component 3: All 5 target cells unlocked AND sheet protection still not enabled (0.1 pts)
                 This compound check fails on initial (target cells still locked) and passes on golden
  Total: 1.0

  Note: Components 2 and 3 are COMPOUND checks anchored to the task change (cell unlocking),
  ensuring they fail on the initial file where target cells are still locked.
"""

import os
import openpyxl
from openpyxl.utils import get_column_letter

WORKDIR = '/home/user'
TASK_ID = 'calc_cop_protection_003'
SHEET_NAME = 'FinancialModel'
TARGET_CELLS = ['B5', 'B8', 'B11', 'B14', 'B17']


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition gate: load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: check the sheet exists
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found in workbook. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Component 1: Individual target cells are unlocked — 0.12 points each (5 x 0.12 = 0.60 total)
    # In the initial file, all target cells have locked=True. After the task, they must be locked=False.
    # This component FAILS on initial (all locked) and PASSES on golden (all unlocked).
    cells_unlocked = []
    cells_still_locked = []

    try:
        for addr in TARGET_CELLS:
            try:
                cell = ws[addr]
                if cell.protection.locked == False:
                    cells_unlocked.append(addr)
                    print(f"PASS: Component 1 — {addr} is unlocked (locked=False) (0.12 pts)")
                    total_score += 0.12
                else:
                    cells_still_locked.append(addr)
                    print(f"FAIL: Component 1 — {addr} is still locked (locked=True) (0.0 pts for this cell)")
            except Exception as e:
                cells_still_locked.append(addr)
                print(f"ERROR: Component 1 — {addr}: {e}")

        unlocked_count = len(cells_unlocked)
        print(f"  -> {unlocked_count}/5 target cells unlocked, partial score: {unlocked_count * 0.12:.2f}")

    except Exception as e:
        print(f"ERROR: Component 1 (outer) — {e}")

    # Component 2: ALL 5 target cells unlocked AND no non-target cells accidentally unlocked (0.3 pts)
    # This is a compound check:
    #   - The first condition (all 5 unlocked) FAILS on initial => whole compound check fails on initial
    #   - On golden: all 5 unlocked AND no other cells unlocked => PASS
    # This ensures the agent only unlocked EXACTLY the 5 specified cells.
    try:
        all_five_unlocked = (len(cells_unlocked) == 5 and len(cells_still_locked) == 0)

        if all_five_unlocked:
            # Now check that no non-target cells were accidentally unlocked
            accidentally_unlocked = []
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    addr = f'{get_column_letter(col)}{row}'
                    if addr not in TARGET_CELLS:
                        cell = ws.cell(row=row, column=col)
                        if cell.protection.locked == False:
                            accidentally_unlocked.append(addr)

            if not accidentally_unlocked:
                print(f"PASS: Component 2 — All 5 target cells unlocked AND no other cells accidentally unlocked (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — All 5 target cells unlocked BUT {len(accidentally_unlocked)} non-target cells accidentally unlocked: {accidentally_unlocked[:10]}{'...' if len(accidentally_unlocked) > 10 else ''} (0.0 pts)")
        else:
            print(f"FAIL: Component 2 — Not all 5 target cells are unlocked (only {len(cells_unlocked)}/5); compound check fails (0.0 pts)")

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: ALL 5 target cells unlocked AND sheet protection still NOT enabled (0.1 pts)
    # This is a compound check:
    #   - The first condition (all 5 unlocked) FAILS on initial => whole compound check fails on initial
    #   - On golden: all 5 unlocked AND no sheet protection => PASS
    # Task explicitly states "Sheet protection should NOT be enabled in this task"
    try:
        all_five_unlocked = (len(cells_unlocked) == 5 and len(cells_still_locked) == 0)

        if all_five_unlocked:
            sheet_protected = ws.protection.sheet
            if not sheet_protected:
                print(f"PASS: Component 3 — All 5 target cells unlocked AND sheet protection is NOT enabled (0.1 pts)")
                total_score += 0.1
            else:
                print(f"FAIL: Component 3 — Sheet protection is enabled (should remain disabled) (0.0 pts)")
        else:
            print(f"FAIL: Component 3 — Not all 5 target cells are unlocked; compound check fails (0.0 pts)")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
