"""
Initial Setup: Marathon tracking spreadsheet and empty output document
Task ID: osworld_multi_apps_book_reading_rate_013
Domain: libreoffice_calc (multi-app: Calc + Writer)

Creates:
  - /home/user/marathons_2023.xlsx  : spreadsheet with race names, locations, empty winning times
  - /home/user/Desktop/fastest_marathon.docx : empty Writer document for agent to fill
"""

import os
import shlex
import subprocess
import time
import openpyxl
from docx import Document

WORKDIR = '/home/user'
DESKTOP = '/home/user/Desktop'
TASK_ID = 'osworld_multi_apps_book_reading_rate_013'
SPREADSHEET = f'{WORKDIR}/marathons_2023.xlsx'
DOCX_PATH = f'{DESKTOP}/fastest_marathon.docx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    # Ensure Desktop directory exists
    os.makedirs(DESKTOP, exist_ok=True)

    # --- Create marathons_2023.xlsx ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Marathons 2023'

    # Headers
    headers = ['Race Name', 'Location', "Men's Winning Time"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Race data — Men's Winning Time is intentionally left empty (agent must fill it)
    races = [
        ('Boston Marathon 2023',  'Boston, MA, USA',     None),
        ('London Marathon 2023',  'London, UK',          None),
        ('Berlin Marathon 2023',  'Berlin, Germany',     None),
        ('Tokyo Marathon 2023',   'Tokyo, Japan',        None),
        ('Chicago Marathon 2023', 'Chicago, IL, USA',    None),
    ]
    for r, (name, location, winning_time) in enumerate(races, 2):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=location)
        ws.cell(row=r, column=3, value=winning_time)  # None = empty cell

    # Set column widths for readability
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 22

    wb.save(SPREADSHEET)
    print(f'Spreadsheet created: {SPREADSHEET}')

    # --- Create empty fastest_marathon.docx on Desktop ---
    doc = Document()
    # Empty document — agent should write the race name here
    doc.save(DOCX_PATH)
    print(f'Empty document created: {DOCX_PATH}')

    # --- GUI-ready startup ---
    # Open the spreadsheet in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{SPREADSHEET}"', delay_sec=2.0)
    # Also open the empty docx so agent can see it is ready
    launch_gui(f'libreoffice --writer "{DOCX_PATH}"', delay_sec=2.0)

    print('GUI_READY: launched LibreOffice Calc and Writer with DISPLAY=:0')


create_initial()
