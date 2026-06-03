"""
Initial Setup: Freeze the first row in the 'Sales Data' sheet
Task ID: calc_sht_freeze_row_001
Domain: libreoffice_calc

Creates a spreadsheet with one sheet 'Sales Data' containing:
- Row 1 headers: Date, Product, Region, Units, Revenue
- 200 rows of realistic sales data (rows 2-201)
- NO freeze panes applied (the agent's task is to add them)
"""

import os
import random
from datetime import date, timedelta
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sht_freeze_row_001'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sales Data'

    # Headers
    headers = ['Date', 'Product', 'Region', 'Units', 'Revenue']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Seed for reproducibility
    random.seed(42)

    products = [
        'Wireless Headphones', 'USB-C Hub', 'Mechanical Keyboard',
        'Monitor Stand', 'Webcam Pro', 'Laptop Sleeve', 'Mouse Pad XL',
        'Portable Charger', 'Desk Lamp LED', 'Cable Organizer'
    ]
    regions = ['North America', 'Europe', 'Asia Pacific', 'Latin America', 'Middle East']

    start_date = date(2024, 1, 2)

    # 200 rows of sales data
    for i in range(200):
        row_num = i + 2
        sale_date = start_date + timedelta(days=random.randint(0, 364))
        product = random.choice(products)
        region = random.choice(regions)
        units = random.randint(1, 150)
        unit_price = round(random.uniform(15.0, 299.0), 2)
        revenue = round(units * unit_price, 2)

        ws.cell(row=row_num, column=1, value=sale_date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=2, value=product)
        ws.cell(row=row_num, column=3, value=region)
        ws.cell(row=row_num, column=4, value=units)
        ws.cell(row=row_num, column=5, value=revenue)

    # NO freeze panes — the agent must add them
    # ws.freeze_panes is left as None (default)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Sales Data')
    print(f'  Rows: 1 header + 200 data rows')
    print(f'  Freeze panes: None (not set)')

create_initial()
