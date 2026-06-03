"""
Initial Setup: Set print range to Selected Sheets and choose specific sheets for printing
Task ID: calc_gsi_093
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_093'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # Style helpers
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    def style_headers(ws, headers, row=1):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def write_data(ws, data, start_row=2):
        for r, row_data in enumerate(data, start_row):
            for c, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = thin_border

    # --- Sheet 1: Executive Summary ---
    ws1 = wb.active
    ws1.title = 'Executive Summary'
    style_headers(ws1, ['Metric', 'Q1 2025', 'Q2 2025', 'Q3 2025', 'Q4 2025', 'FY 2025'])
    data1 = [
        ['Total Revenue ($M)', 12.4, 13.8, 15.2, 16.1, 57.5],
        ['Gross Margin (%)', 62.3, 63.1, 64.5, 65.2, 63.8],
        ['Operating Income ($M)', 2.1, 2.5, 3.0, 3.4, 11.0],
        ['Net Income ($M)', 1.6, 1.9, 2.3, 2.7, 8.5],
        ['EPS ($)', 0.32, 0.38, 0.46, 0.54, 1.70],
        ['Headcount', 234, 241, 258, 267, 267],
        ['Customer Count', 1842, 1923, 2087, 2201, 2201],
        ['NPS Score', 72, 74, 76, 78, 75],
        ['ARR ($M)', 48.2, 51.6, 55.8, 59.4, 59.4],
        ['Churn Rate (%)', 3.2, 2.9, 2.7, 2.5, 2.8],
        ['CAC ($)', 2450, 2380, 2290, 2150, 2318],
        ['LTV/CAC Ratio', 4.2, 4.5, 4.8, 5.1, 4.7],
    ]
    write_data(ws1, data1)
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws1.column_dimensions[col].width = 14
    ws1.column_dimensions['A'].width = 24

    # --- Sheet 2: Financial Results ---
    ws2 = wb.create_sheet('Financial Results')
    style_headers(ws2, ['Category', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                        'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Total'])
    fin_data = [
        ['Product Revenue', 980, 1020, 1150, 1080, 1190, 1230, 1280, 1310, 1350, 1290, 1380, 1420, 14680],
        ['Service Revenue', 320, 340, 360, 350, 370, 390, 400, 410, 430, 420, 440, 460, 4690],
        ['License Revenue', 150, 160, 170, 180, 190, 200, 210, 220, 230, 240, 250, 260, 2460],
        ['COGS - Product', -340, -355, -400, -375, -415, -430, -445, -455, -470, -450, -480, -495, -5110],
        ['COGS - Service', -160, -170, -180, -175, -185, -195, -200, -205, -215, -210, -220, -230, -2345],
        ['Gross Profit', 950, 995, 1100, 1060, 1150, 1195, 1245, 1280, 1325, 1290, 1370, 1415, 14375],
        ['Sales & Marketing', -280, -290, -310, -300, -320, -330, -340, -350, -360, -350, -370, -380, -3980],
        ['R&D Expense', -210, -215, -220, -225, -230, -235, -240, -245, -250, -255, -260, -265, -2850],
        ['G&A Expense', -120, -125, -130, -128, -132, -135, -138, -140, -142, -145, -148, -150, -1633],
        ['Operating Income', 340, 365, 440, 407, 468, 495, 527, 545, 573, 540, 592, 620, 5912],
        ['Interest Income', 12, 13, 14, 13, 14, 15, 15, 16, 16, 17, 17, 18, 180],
        ['Tax Expense', -88, -95, -114, -105, -121, -128, -136, -140, -147, -139, -152, -159, -1524],
    ]
    write_data(ws2, fin_data)
    ws2.column_dimensions['A'].width = 22
    for col_letter in ['B','C','D','E','F','G','H','I','J','K','L','M','N']:
        ws2.column_dimensions[col_letter].width = 12

    # --- Sheet 3: KPIs ---
    ws3 = wb.create_sheet('KPIs')
    style_headers(ws3, ['KPI Name', 'Target', 'Actual', 'Status', 'Trend', 'Owner'])
    kpi_data = [
        ['Monthly Recurring Revenue', '$4.95M', '$5.12M', 'On Track', 'Up', 'Sarah Chen'],
        ['Customer Acquisition Cost', '$2,200', '$2,150', 'On Track', 'Down', 'Marcus Johnson'],
        ['Customer Lifetime Value', '$10,500', '$10,965', 'Exceeding', 'Up', 'Sarah Chen'],
        ['Monthly Active Users', '45,000', '47,832', 'On Track', 'Up', 'Priya Patel'],
        ['Support Ticket Resolution', '< 4 hours', '3.2 hours', 'On Track', 'Stable', 'David Kim'],
        ['Feature Adoption Rate', '60%', '58%', 'At Risk', 'Down', 'Lisa Wang'],
        ['Employee Satisfaction', '4.2/5.0', '4.3/5.0', 'On Track', 'Up', 'Rachel Torres'],
        ['Code Deploy Frequency', '12/week', '14/week', 'Exceeding', 'Up', 'James Liu'],
        ['System Uptime', '99.95%', '99.97%', 'On Track', 'Stable', 'Omar Hassan'],
        ['Revenue Per Employee', '$215K', '$222K', 'On Track', 'Up', 'Sarah Chen'],
        ['Pipeline Coverage', '3.5x', '3.8x', 'On Track', 'Up', 'Marcus Johnson'],
        ['Win Rate', '28%', '31%', 'Exceeding', 'Up', 'Marcus Johnson'],
    ]
    write_data(ws3, kpi_data)
    ws3.column_dimensions['A'].width = 28
    ws3.column_dimensions['B'].width = 14
    ws3.column_dimensions['C'].width = 14
    ws3.column_dimensions['D'].width = 14
    ws3.column_dimensions['E'].width = 10
    ws3.column_dimensions['F'].width = 18

    # --- Sheet 4: Revenue Breakdown ---
    ws4 = wb.create_sheet('Revenue Breakdown')
    style_headers(ws4, ['Product Line', 'Region', 'Q1', 'Q2', 'Q3', 'Q4', 'Annual'])
    rev_data = [
        ['Enterprise Suite', 'North America', 2100, 2250, 2400, 2550, 9300],
        ['Enterprise Suite', 'EMEA', 1400, 1500, 1600, 1700, 6200],
        ['Enterprise Suite', 'APAC', 800, 860, 920, 980, 3560],
        ['SMB Platform', 'North America', 1600, 1720, 1840, 1960, 7120],
        ['SMB Platform', 'EMEA', 950, 1020, 1090, 1160, 4220],
        ['SMB Platform', 'APAC', 620, 665, 710, 755, 2750],
        ['Developer Tools', 'North America', 480, 515, 550, 585, 2130],
        ['Developer Tools', 'EMEA', 320, 345, 370, 395, 1430],
        ['Developer Tools', 'APAC', 210, 225, 240, 255, 930],
        ['API Services', 'North America', 380, 410, 440, 470, 1700],
        ['API Services', 'EMEA', 250, 270, 290, 310, 1120],
        ['API Services', 'APAC', 170, 185, 200, 215, 770],
    ]
    write_data(ws4, rev_data)
    ws4.column_dimensions['A'].width = 20
    ws4.column_dimensions['B'].width = 18

    # --- Sheet 5: Headcount ---
    ws5 = wb.create_sheet('Headcount')
    style_headers(ws5, ['Department', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                        'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'])
    hc_data = [
        ['Engineering', 82, 83, 85, 87, 89, 91, 93, 95, 97, 99, 101, 103],
        ['Sales', 38, 38, 39, 40, 41, 42, 43, 44, 45, 45, 46, 47],
        ['Marketing', 22, 22, 23, 23, 24, 24, 25, 25, 26, 26, 27, 27],
        ['Customer Success', 28, 28, 29, 30, 30, 31, 32, 32, 33, 34, 34, 35],
        ['Product', 18, 18, 19, 19, 20, 20, 21, 21, 22, 22, 23, 23],
        ['Finance', 12, 12, 12, 12, 13, 13, 13, 13, 14, 14, 14, 14],
        ['HR & Admin', 10, 10, 10, 10, 11, 11, 11, 11, 11, 12, 12, 12],
        ['Legal', 5, 5, 5, 5, 5, 6, 6, 6, 6, 6, 6, 6],
        ['IT Operations', 8, 8, 8, 9, 9, 9, 9, 10, 10, 10, 10, 10],
    ]
    write_data(ws5, hc_data)
    ws5.column_dimensions['A'].width = 22

    # --- Sheet 6: Marketing Spend ---
    ws6 = wb.create_sheet('Marketing Spend')
    style_headers(ws6, ['Channel', 'Budget ($K)', 'Actual ($K)', 'Variance ($K)', 'ROI (%)', 'Leads Generated'])
    mkt_data = [
        ['Google Ads', 180, 175, 5, 320, 4250],
        ['LinkedIn Ads', 120, 128, -8, 280, 2180],
        ['Content Marketing', 85, 82, 3, 410, 3620],
        ['Email Campaigns', 45, 43, 2, 520, 5840],
        ['Trade Shows', 200, 215, -15, 180, 1420],
        ['Webinars', 35, 33, 2, 450, 2890],
        ['Social Media', 60, 58, 2, 290, 2340],
        ['PR & Analyst', 95, 92, 3, 150, 890],
        ['Partner Marketing', 75, 71, 4, 340, 1960],
        ['SEO', 40, 38, 2, 580, 6210],
    ]
    write_data(ws6, mkt_data)
    ws6.column_dimensions['A'].width = 22
    ws6.column_dimensions['F'].width = 18

    # --- Sheet 7: Operational Costs ---
    ws7 = wb.create_sheet('Operational Costs')
    style_headers(ws7, ['Cost Center', 'Monthly ($K)', 'Annual ($K)', 'YoY Change (%)', 'Notes'])
    ops_data = [
        ['Cloud Infrastructure (AWS)', 142, 1704, 18.5, 'Migration to reserved instances in Q3'],
        ['Office Lease - SF HQ', 85, 1020, 0.0, 'Lease renewal due Dec 2025'],
        ['Office Lease - Austin', 32, 384, 5.2, 'Expanded to 2nd floor in Feb'],
        ['Software Licenses', 28, 336, 12.0, 'Added Salesforce Enterprise'],
        ['Insurance', 15, 180, 3.5, 'Standard annual increase'],
        ['Utilities & Telecom', 8, 96, 2.1, 'Includes fiber upgrade'],
        ['Travel & Entertainment', 22, 264, -8.0, 'Reduced post-COVID normalization'],
        ['Professional Services', 35, 420, 15.0, 'External audit + legal counsel'],
        ['Equipment & Hardware', 18, 216, -5.0, 'Shifted to BYOD policy'],
        ['Miscellaneous', 6, 72, 1.5, 'Office supplies, subscriptions'],
    ]
    write_data(ws7, ops_data)
    ws7.column_dimensions['A'].width = 30
    ws7.column_dimensions['E'].width = 40

    # --- Sheet 8: Customer Metrics ---
    ws8 = wb.create_sheet('Customer Metrics')
    style_headers(ws8, ['Segment', 'Customers', 'Avg Deal Size ($)', 'Retention (%)',
                        'Expansion Rate (%)', 'Support Tickets/Mo'])
    cust_data = [
        ['Enterprise (>500 emp)', 42, 185000, 96.2, 22.5, 128],
        ['Mid-Market (100-500)', 186, 52000, 92.8, 18.3, 342],
        ['SMB (10-100)', 824, 12500, 88.5, 12.1, 890],
        ['Startup (<10)', 1149, 3200, 82.3, 8.4, 1245],
        ['Government', 18, 220000, 98.1, 5.2, 45],
        ['Education', 34, 28000, 94.5, 10.8, 78],
        ['Non-Profit', 22, 8500, 91.2, 7.3, 52],
        ['Healthcare', 56, 95000, 95.8, 19.7, 165],
        ['Financial Services', 38, 142000, 97.3, 24.1, 112],
        ['Retail/E-Commerce', 72, 38000, 89.4, 14.6, 198],
    ]
    write_data(ws8, cust_data)
    ws8.column_dimensions['A'].width = 26
    ws8.column_dimensions['C'].width = 18
    ws8.column_dimensions['F'].width = 20

    # --- Sheet 9: R&D Pipeline ---
    ws9 = wb.create_sheet('R&D Pipeline')
    style_headers(ws9, ['Project', 'Lead', 'Phase', 'Start Date', 'Target Release',
                        'Budget ($K)', 'Status'])
    rd_data = [
        ['AI Assistant v2', 'James Liu', 'Development', '2025-01-15', '2025-06-30', 450, 'On Track'],
        ['Mobile App Redesign', 'Priya Patel', 'Design', '2025-02-01', '2025-08-15', 280, 'On Track'],
        ['API Gateway v3', 'Omar Hassan', 'Testing', '2024-10-01', '2025-04-30', 320, 'At Risk'],
        ['Analytics Dashboard', 'Lisa Wang', 'Development', '2025-01-20', '2025-07-15', 190, 'On Track'],
        ['SSO Integration', 'David Kim', 'Planning', '2025-03-01', '2025-09-30', 120, 'Not Started'],
        ['Performance Optimizer', 'James Liu', 'Research', '2025-02-15', '2025-12-31', 380, 'On Track'],
        ['Compliance Module', 'Rachel Torres', 'Development', '2024-11-15', '2025-05-31', 250, 'Delayed'],
        ['Customer Portal v2', 'Priya Patel', 'Testing', '2024-09-01', '2025-04-15', 210, 'On Track'],
        ['Data Export Tool', 'Omar Hassan', 'Development', '2025-01-05', '2025-05-30', 95, 'On Track'],
        ['Workflow Automation', 'Lisa Wang', 'Planning', '2025-03-15', '2025-10-31', 340, 'Not Started'],
    ]
    write_data(ws9, rd_data)
    ws9.column_dimensions['A'].width = 24
    ws9.column_dimensions['B'].width = 16
    ws9.column_dimensions['C'].width = 14
    ws9.column_dimensions['D'].width = 14
    ws9.column_dimensions['E'].width = 16
    ws9.column_dimensions['G'].width = 14

    # --- Sheet 10: Board Notes ---
    ws10 = wb.create_sheet('Board Notes')
    style_headers(ws10, ['Topic', 'Discussion Points', 'Action Items', 'Owner', 'Due Date'])
    board_data = [
        ['Q4 Financial Review', 'Revenue exceeded target by 4.2%. Margins improved due to cost optimization.', 'Prepare investor update deck', 'Sarah Chen', '2025-02-15'],
        ['Product Roadmap', 'AI features prioritized for H1. Mobile redesign moved to Q2.', 'Finalize H1 feature priorities', 'Priya Patel', '2025-02-10'],
        ['Hiring Plan', 'Approved 45 new positions. Focus on engineering and sales.', 'Submit req forms to HR', 'Rachel Torres', '2025-02-20'],
        ['Market Expansion', 'APAC growth strong at 22% YoY. Consider Japan office.', 'Research Japan market entry costs', 'Marcus Johnson', '2025-03-15'],
        ['Customer Success', 'NPS improved to 78. Enterprise retention at 96.2%.', 'Develop case study program', 'David Kim', '2025-02-28'],
        ['Security & Compliance', 'SOC2 Type II audit scheduled for Q2. GDPR update needed.', 'Engage compliance consultant', 'Omar Hassan', '2025-03-01'],
        ['Partnership Strategy', 'AWS partnership generating 15% of pipeline. Explore Azure.', 'Draft Azure partnership proposal', 'Marcus Johnson', '2025-03-10'],
        ['Budget Reallocation', 'Shift $200K from trade shows to digital marketing based on ROI.', 'Update marketing budget model', 'Lisa Wang', '2025-02-25'],
    ]
    write_data(ws10, board_data)
    ws10.column_dimensions['A'].width = 24
    ws10.column_dimensions['B'].width = 60
    ws10.column_dimensions['C'].width = 35
    ws10.column_dimensions['D'].width = 18
    ws10.column_dimensions['E'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the workbook in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
