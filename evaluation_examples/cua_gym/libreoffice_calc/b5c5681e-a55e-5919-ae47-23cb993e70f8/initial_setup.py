"""
Initial Setup: Set up sheet events - assign macro 'LogAccess' to 'Secure Data' sheet activate event.
Task ID: calc_ps_081
Domain: libreoffice_calc

Initial state:
- Workbook with multiple sheets including 'Secure Data'
- A macro 'LogAccess' exists in the Standard macro library
- NO sheet events are configured (task requires the user to do that)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_081'
XLSX_TEMP = f'{WORKDIR}/{TASK_ID}_temp.xlsx'
OUTPUT = f'{WORKDIR}/{TASK_ID}.ods'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Secure Data ---
    ws1 = wb.active
    ws1.title = 'Secure Data'

    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    headers = ['Employee ID', 'Full Name', 'Clearance Level', 'Department', 'Access Code', 'Last Access Date']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    col_widths = {'A': 14, 'B': 22, 'C': 16, 'D': 20, 'E': 14, 'F': 18}
    for col_letter, width in col_widths.items():
        ws1.column_dimensions[col_letter].width = width

    data = [
        ['EMP-1001', 'Sarah Chen', 'Top Secret', 'Cybersecurity', 'AX-7742', '2025-11-20'],
        ['EMP-1002', 'Marcus Johnson', 'Secret', 'Research & Dev', 'BK-3318', '2025-12-01'],
        ['EMP-1003', 'Elena Vasquez', 'Top Secret', 'Intelligence', 'CZ-9905', '2025-11-28'],
        ['EMP-1004', 'James O\'Brien', 'Confidential', 'Finance', 'DW-5561', '2025-10-15'],
        ['EMP-1005', 'Aisha Patel', 'Top Secret', 'Cybersecurity', 'EQ-2234', '2025-12-03'],
        ['EMP-1006', 'Robert Kim', 'Secret', 'Operations', 'FT-8877', '2025-11-10'],
        ['EMP-1007', 'Laura Martinez', 'Confidential', 'Human Resources', 'GN-1149', '2025-09-22'],
        ['EMP-1008', 'David Thompson', 'Top Secret', 'Intelligence', 'HP-6623', '2025-12-05'],
        ['EMP-1009', 'Mei Lin Wang', 'Secret', 'Research & Dev', 'JR-4456', '2025-11-30'],
        ['EMP-1010', 'Carlos Rivera', 'Confidential', 'Legal', 'KS-7790', '2025-10-28'],
        ['EMP-1011', 'Natasha Ivanova', 'Top Secret', 'Cybersecurity', 'LU-3312', '2025-12-04'],
        ['EMP-1012', 'William Hayes', 'Secret', 'Operations', 'MV-5548', '2025-11-18'],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 3:
                if val == 'Top Secret':
                    cell.font = Font(color='FF0000', bold=True)
                elif val == 'Secret':
                    cell.font = Font(color='FF8C00')

    # --- Sheet 2: Access Log ---
    ws2 = wb.create_sheet('Access Log')
    log_headers = ['Timestamp', 'Employee ID', 'Action', 'IP Address', 'Status']
    for col, h in enumerate(log_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFD9E2F3', end_color='FFD9E2F3', fill_type='solid')

    log_data = [
        ['2025-12-01 08:15:22', 'EMP-1001', 'View', '192.168.1.101', 'Granted'],
        ['2025-12-01 09:30:45', 'EMP-1003', 'Export', '192.168.1.105', 'Granted'],
        ['2025-12-01 10:12:08', 'EMP-1007', 'View', '192.168.1.142', 'Denied'],
        ['2025-12-01 11:45:33', 'EMP-1005', 'Modify', '192.168.1.118', 'Granted'],
        ['2025-12-01 13:20:17', 'EMP-1002', 'View', '192.168.1.110', 'Granted'],
        ['2025-12-01 14:55:44', 'EMP-1010', 'Export', '192.168.1.155', 'Denied'],
        ['2025-12-02 08:05:11', 'EMP-1008', 'View', '192.168.1.130', 'Granted'],
        ['2025-12-02 09:40:29', 'EMP-1011', 'Modify', '192.168.1.162', 'Granted'],
    ]
    for r, row_data in enumerate(log_data, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    col_widths2 = {'A': 22, 'B': 14, 'C': 10, 'D': 16, 'E': 10}
    for col_letter, width in col_widths2.items():
        ws2.column_dimensions[col_letter].width = width

    # --- Sheet 3: Summary ---
    ws3 = wb.create_sheet('Summary')
    ws3['A1'] = 'Security Clearance Summary'
    ws3['A1'].font = Font(size=14, bold=True)
    ws3['A3'] = 'Clearance Level'
    ws3['B3'] = 'Count'
    ws3['A4'] = 'Top Secret'
    ws3['B4'] = 5
    ws3['A5'] = 'Secret'
    ws3['B5'] = 4
    ws3['A6'] = 'Confidential'
    ws3['B6'] = 3
    ws3['A8'] = 'Total Employees'
    ws3['B8'] = 12
    ws3['A8'].font = Font(bold=True)
    ws3['B8'].font = Font(bold=True)

    wb.save(XLSX_TEMP)
    print(f'Temp xlsx created: {XLSX_TEMP}')

    # --- Install the LogAccess macro ---
    macro_dir = os.path.expanduser('~/.config/libreoffice/4/user/basic/Standard')
    os.makedirs(macro_dir, exist_ok=True)

    macro_content = """Sub LogAccess()
    ' This macro logs when the Secure Data sheet is accessed
    Dim oSheet As Object
    Dim oDoc As Object
    Dim oLogSheet As Object
    Dim lastRow As Long

    oDoc = ThisComponent

    ' Find Access Log sheet
    If oDoc.Sheets.hasByName("Access Log") Then
        oLogSheet = oDoc.Sheets.getByName("Access Log")
        lastRow = 0
        Do While oLogSheet.getCellByPosition(0, lastRow).getString() <> ""
            lastRow = lastRow + 1
        Loop

        oLogSheet.getCellByPosition(0, lastRow).setString(Format(Now(), "yyyy-mm-dd hh:mm:ss"))
        oLogSheet.getCellByPosition(1, lastRow).setString("SYSTEM")
        oLogSheet.getCellByPosition(2, lastRow).setString("Sheet Activated")
        oLogSheet.getCellByPosition(3, lastRow).setString("localhost")
        oLogSheet.getCellByPosition(4, lastRow).setString("Logged")
    End If
End Sub
"""

    module_path = os.path.join(macro_dir, 'LogAccess.xba')
    with open(module_path, 'w') as f:
        f.write('<?xml version="1.0" encoding="UTF-8"?>\n')
        f.write('<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">\n')
        f.write('<script:module xmlns:script="http://openoffice.org/2000/script" script:name="LogAccess" script:language="StarBasic">')
        f.write(macro_content)
        f.write('</script:module>\n')

    module1_path = os.path.join(macro_dir, 'Module1.xba')
    if not os.path.exists(module1_path):
        with open(module1_path, 'w') as f:
            f.write('<?xml version="1.0" encoding="UTF-8"?>\n')
            f.write('<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">\n')
            f.write('<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">\n')
            f.write('REM  *****  BASIC  *****\n\n')
            f.write('Sub Main\n\nEnd Sub\n')
            f.write('</script:module>\n')

    lib_path = os.path.join(macro_dir, 'script-lb.xml')
    with open(lib_path, 'w') as f:
        f.write('<?xml version="1.0" encoding="UTF-8"?>\n')
        f.write('<!DOCTYPE library:library PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "library.dtd">\n')
        f.write('<library:library xmlns:library="http://openoffice.org/2000/library" library:name="Standard" library:readonly="false" library:passwordprotected="false">\n')
        f.write(' <library:element library:name="LogAccess"/>\n')
        f.write(' <library:element library:name="Module1"/>\n')
        f.write('</library:library>\n')

    print(f'Macro LogAccess installed')

    # Convert xlsx to ods using LibreOffice headless
    env = os.environ.copy()
    env["DISPLAY"] = ":0"

    # Kill any running LibreOffice first
    subprocess.run(['pkill', '-f', 'soffice'], capture_output=True)
    time.sleep(2)

    result = subprocess.run(
        ['libreoffice', '--headless', '--convert-to', 'ods', '--outdir', WORKDIR, XLSX_TEMP],
        capture_output=True, text=True, timeout=60,
        env=env
    )
    print(f'Convert stdout: {result.stdout}')
    if result.stderr:
        print(f'Convert stderr: {result.stderr}')

    # The converted file will be calc_ps_081_temp.ods, rename to calc_ps_081.ods
    temp_ods = f'{WORKDIR}/{TASK_ID}_temp.ods'
    if os.path.exists(temp_ods):
        os.rename(temp_ods, OUTPUT)
        print(f'Renamed {temp_ods} -> {OUTPUT}')

    # Clean up temp xlsx
    if os.path.exists(XLSX_TEMP):
        os.remove(XLSX_TEMP)

    # Verify output exists
    if os.path.exists(OUTPUT):
        print(f'Initial ODS file created: {OUTPUT}')
    else:
        print(f'ERROR: ODS file not found at {OUTPUT}')

    # Open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
