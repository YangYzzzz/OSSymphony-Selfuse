"""
Initial Setup: Time tracking spreadsheet for WEEKDAY-based conditional formatting task
Task ID: osworld_calc_conditional_format_weekday_006
Domain: libreoffice_calc

Creates a spreadsheet with daily work hours tracking data.
No conditional formatting applied (agent will add it).
"""

import os
import shlex
import subprocess
import time
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_conditional_format_weekday_006'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Time Tracking ---
    ws = wb.active
    ws.title = "Time Tracking"

    # Headers
    headers = ["Date", "Employee", "Project", "Hours"]
    header_font = Font(name="Calibri", bold=True, size=11)
    header_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Realistic time tracking data
    # Mix of weekday and weekend dates, some overtime (hours > 8)
    # Using dates from March 2025 (mix of weekdays Mon-Fri and weekends Sat-Sun)
    data = [
        # date, employee, project, hours
        (date(2025, 3, 3),  "Sarah Chen",       "Website Redesign",     8.0),   # Monday
        (date(2025, 3, 4),  "Marcus Johnson",   "Client Portal",        9.5),   # Tuesday (overtime)
        (date(2025, 3, 5),  "Emily Rodriguez",  "Data Migration",       7.5),   # Wednesday
        (date(2025, 3, 6),  "James Park",       "Mobile App",           8.0),   # Thursday
        (date(2025, 3, 7),  "Sarah Chen",       "Website Redesign",     10.0),  # Friday (overtime)
        (date(2025, 3, 8),  "Marcus Johnson",   "Client Portal",        4.0),   # Saturday (weekend)
        (date(2025, 3, 9),  "Emily Rodriguez",  "Data Migration",       3.5),   # Sunday (weekend)
        (date(2025, 3, 10), "James Park",       "Mobile App",           8.0),   # Monday
        (date(2025, 3, 11), "Sarah Chen",       "Website Redesign",     9.0),   # Tuesday (overtime)
        (date(2025, 3, 12), "Marcus Johnson",   "Client Portal",        7.0),   # Wednesday
        (date(2025, 3, 13), "Emily Rodriguez",  "Data Migration",       8.5),   # Thursday (overtime)
        (date(2025, 3, 14), "James Park",       "Mobile App",           6.0),   # Friday
        (date(2025, 3, 15), "Sarah Chen",       "Website Redesign",     5.0),   # Saturday (weekend)
        (date(2025, 3, 16), "Marcus Johnson",   "Client Portal",        2.5),   # Sunday (weekend)
        (date(2025, 3, 17), "Emily Rodriguez",  "Data Migration",       8.0),   # Monday
        (date(2025, 3, 18), "James Park",       "Mobile App",           11.0),  # Tuesday (overtime)
        (date(2025, 3, 19), "Sarah Chen",       "Website Redesign",     7.5),   # Wednesday
        (date(2025, 3, 20), "Marcus Johnson",   "Client Portal",        8.0),   # Thursday
        (date(2025, 3, 21), "Emily Rodriguez",  "Data Migration",       9.5),   # Friday (overtime)
        (date(2025, 3, 22), "James Park",       "Mobile App",           6.0),   # Saturday (weekend)
        (date(2025, 3, 23), "Sarah Chen",       "Website Redesign",     4.0),   # Sunday (weekend)
        (date(2025, 3, 24), "Marcus Johnson",   "Client Portal",        8.0),   # Monday
    ]

    for r, (row_date, employee, project, hours) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_date).number_format = 'yyyy-mm-dd'
        ws.cell(row=r, column=2, value=employee)
        ws.cell(row=r, column=3, value=project)
        ws.cell(row=r, column=4, value=hours)

    # Set column widths for readability
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 10

    # Freeze header row
    ws.freeze_panes = "A2"

    # NO conditional formatting — agent will add this

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
