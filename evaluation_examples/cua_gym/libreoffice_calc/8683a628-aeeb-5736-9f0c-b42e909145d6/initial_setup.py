"""
Initial Setup: Customer segmentation table with account data
Task ID: calc_sales_031
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_031'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Accounts'

    # --- Headers ---
    headers_main = {1: 'Account', 2: 'Annual Revenue', 3: 'Segment'}
    for col, h in headers_main.items():
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Summary headers in E1:F1
    ws.cell(row=1, column=5, value='Segment').font = Font(bold=True)
    ws.cell(row=1, column=6, value='Count').font = Font(bold=True)
    ws['E1'].alignment = Alignment(horizontal='center')
    ws['F1'].alignment = Alignment(horizontal='center')

    # --- Account Data (A2:B9) ---
    accounts = [
        ('Acme Corp', 750000),
        ('Beta Inc', 120000),
        ('Gamma LLC', 45000),
        ('Delta Co', 320000),
        ('Epsilon Ltd', 890000),
        ('Zeta Corp', 95000),
        ('Eta Inc', 210000),
        ('Theta Co', 500001),
    ]
    for r, (name, revenue) in enumerate(accounts, 2):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=revenue)
        ws.cell(row=r, column=2).number_format = '$#,##0'

    # C2:C9 intentionally left EMPTY (agent must fill with IF formulas)

    # --- Segment labels (E2:E4) ---
    ws.cell(row=2, column=5, value='Enterprise')
    ws.cell(row=3, column=5, value='Mid-Market')
    ws.cell(row=4, column=5, value='SMB')

    # F2:F4 intentionally left EMPTY (agent must fill with COUNTIF formulas)

    # --- Column widths ---
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 3
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
