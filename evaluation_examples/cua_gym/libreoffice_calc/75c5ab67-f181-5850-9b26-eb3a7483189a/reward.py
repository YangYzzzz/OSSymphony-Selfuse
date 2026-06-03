"""
Reward Script: Calculate compa-ratio for each employee (Salary / Band Midpoint)
Task ID: calc_hr_042
Domain: libreoffice_calc
Scoring:
  Component 1 (0.50): D2:D6 contain correct division formulas =Bn/Cn
  Component 2 (0.30): D2:D6 are formatted as percentage
  Component 3 (0.20): Computed values match expected compa-ratios
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_042'

# Expected formulas (case-insensitive, whitespace-normalized)
EXPECTED_FORMULAS = {
    2: '=B2/C2',
    3: '=B3/C3',
    4: '=B4/C4',
    5: '=B5/C5',
    6: '=B6/C6',
}

# Expected computed values (Salary / Midpoint)
EXPECTED_VALUES = {
    2: 72000 / 75000,   # 0.96
    3: 85000 / 80000,   # 1.0625
    4: 62000 / 70000,   # 0.885714...
    5: 110000 / 105000, # 1.047619...
    6: 54000 / 55000,   # 0.981818...
}


def normalize_formula(f):
    """Normalize a formula string for comparison."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook (with formulas)
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists
    if 'Equity' not in wb.sheetnames:
        print("FAIL: Sheet 'Equity' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Equity']

    # Component 1: D2:D6 contain correct division formulas (0.50 points)
    # Each correct formula earns 0.10 points
    try:
        formula_score = 0.0
        for row in range(2, 7):
            cell_val = ws.cell(row=row, column=4).value
            expected = EXPECTED_FORMULAS[row]
            if cell_val is not None and normalize_formula(str(cell_val)) == normalize_formula(expected):
                print(f"PASS: D{row} formula correct: {cell_val}")
                formula_score += 0.10
            else:
                print(f"FAIL: D{row} expected formula {expected}, found: {cell_val}")
        if formula_score > 0:
            print(f"PASS: Component 1 — formulas ({formula_score:.2f} of 0.50 pts)")
            total_score += formula_score
        else:
            print("FAIL: Component 1 — no correct formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: D2:D6 formatted as percentage (0.30 points)
    # Each correct format earns 0.06 points
    try:
        fmt_score = 0.0
        for row in range(2, 7):
            cell = ws.cell(row=row, column=4)
            nf = cell.number_format if cell.number_format else ''
            if '%' in nf:
                print(f"PASS: D{row} format is percentage: {nf}")
                fmt_score += 0.06
            else:
                print(f"FAIL: D{row} expected percentage format, found: {nf}")
        if fmt_score > 0:
            print(f"PASS: Component 2 — percentage formatting ({fmt_score:.2f} of 0.30 pts)")
            total_score += fmt_score
        else:
            print("FAIL: Component 2 — no percentage formatting found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Computed values match expected (0.20 points)
    # Load with data_only to get cached values; if not available, compute from B/C columns
    try:
        value_score = 0.0
        # Try data_only first
        wb_data = openpyxl.load_workbook(file_path, data_only=True)
        ws_data = wb_data['Equity']

        for row in range(2, 7):
            d_val = ws_data.cell(row=row, column=4).value
            b_val = ws.cell(row=row, column=2).value
            c_val = ws.cell(row=row, column=3).value
            expected_val = EXPECTED_VALUES[row]

            # Check if data_only returned a computed value
            if d_val is not None and isinstance(d_val, (int, float)):
                if abs(float(d_val) - expected_val) < 0.005:
                    print(f"PASS: D{row} value correct (data_only): {d_val:.4f} ~ {expected_val:.4f}")
                    value_score += 0.04
                else:
                    print(f"FAIL: D{row} value mismatch: {d_val} vs expected {expected_val:.4f}")
            else:
                # data_only returned None (file not opened by Calc yet)
                # Fallback: verify formula is a division of B/C, and B,C have correct values
                formula = ws.cell(row=row, column=4).value
                if (formula is not None and
                    isinstance(b_val, (int, float)) and
                    isinstance(c_val, (int, float)) and
                    normalize_formula(str(formula)) == normalize_formula(EXPECTED_FORMULAS[row])):
                    # Formula correct + source values correct -> result would be correct
                    computed = float(b_val) / float(c_val)
                    if abs(computed - expected_val) < 0.005:
                        print(f"PASS: D{row} formula+sources verify correct result: {computed:.4f}")
                        value_score += 0.04
                    else:
                        print(f"FAIL: D{row} computed {computed:.4f} != expected {expected_val:.4f}")
                else:
                    print(f"FAIL: D{row} no cached value and formula/sources invalid")

        if value_score > 0:
            print(f"PASS: Component 3 — computed values ({value_score:.2f} of 0.20 pts)")
            total_score += value_score
        else:
            print("FAIL: Component 3 — no correct computed values")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
