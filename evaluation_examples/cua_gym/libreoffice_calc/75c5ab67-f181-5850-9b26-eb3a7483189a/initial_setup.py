"""
Initial Setup: Calculate compa-ratio for each employee to assess pay equity
Task ID: calc_hr_042
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_042'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Equity ---
    ws = wb.active
    ws.title = 'Equity'

    # Headers
    headers = ['Employee', 'Salary', 'Band Midpoint', 'Compa-Ratio']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Employee data - realistic HR dataset
    data = [
        ['Alice', 72000, 75000],
        ['Bob', 85000, 80000],
        ['Carol', 62000, 70000],
        ['Dan', 110000, 105000],
        ['Eve', 54000, 55000],
    ]

    data_font = Font(name='Calibri', size=11)
    currency_format = '$#,##0'

    for r, row_data in enumerate(data, 2):
        # Column A: Employee name
        cell_a = ws.cell(row=r, column=1, value=row_data[0])
        cell_a.font = data_font
        cell_a.border = thin_border

        # Column B: Salary
        cell_b = ws.cell(row=r, column=2, value=row_data[1])
        cell_b.font = data_font
        cell_b.number_format = currency_format
        cell_b.border = thin_border

        # Column C: Band Midpoint
        cell_c = ws.cell(row=r, column=3, value=row_data[2])
        cell_c.font = data_font
        cell_c.number_format = currency_format
        cell_c.border = thin_border

        # Column D: Compa-Ratio - LEFT EMPTY (task requires agent to fill this)
        cell_d = ws.cell(row=r, column=4)
        cell_d.border = thin_border

    # Set column widths for readability
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
