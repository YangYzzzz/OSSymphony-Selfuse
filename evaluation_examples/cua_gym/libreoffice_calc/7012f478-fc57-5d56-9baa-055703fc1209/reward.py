"""
Reward Script: Format cells C2:C20 as currency ($#,##0.00) AND right-align them, then center-align the header C1.
Task ID: calc_fmt_number_and_align_combined_064
Domain: libreoffice_calc
Scoring:
  - Component 1: C2:C20 have number format '$#,##0.00' (0.4 pts)
  - Component 2: C2:C20 are right-aligned (0.4 pts)
  - Component 3: C1 is center-aligned (0.2 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_number_and_align_combined_064'
SHEET_NAME = 'Price List'
CURRENCY_FORMAT = '$#,##0.00'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet 'Price List' must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found in workbook. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Component 1: C2:C20 have number format '$#,##0.00' (0.4 points)
    # This FAILS on initial (all 'General') and PASSES on golden (all '$#,##0.00')
    try:
        currency_count = 0
        currency_failures = []
        for row in range(2, 21):
            cell = ws.cell(row=row, column=3)
            fmt = cell.number_format
            if fmt == CURRENCY_FORMAT:
                currency_count += 1
            else:
                currency_failures.append(f"C{row}: {repr(fmt)}")

        if currency_count == 19:
            print(f"PASS: Component 1 — All 19 cells C2:C20 have number format '{CURRENCY_FORMAT}' (0.4 pts)")
            total_score += 0.4
        elif currency_count > 0:
            print(f"PARTIAL: Component 1 — Only {currency_count}/19 cells in C2:C20 have format '{CURRENCY_FORMAT}'. Failures: {currency_failures[:5]}")
        else:
            print(f"FAIL: Component 1 — No cells in C2:C20 have format '{CURRENCY_FORMAT}'. Sample: {currency_failures[:3]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: C2:C20 are right-aligned (0.4 points)
    # This FAILS on initial (all 'left' or 'general') and PASSES on golden (all 'right')
    try:
        right_count = 0
        align_failures = []
        for row in range(2, 21):
            cell = ws.cell(row=row, column=3)
            align = cell.alignment.horizontal
            if align == 'right':
                right_count += 1
            else:
                align_failures.append(f"C{row}: {repr(align)}")

        if right_count == 19:
            print(f"PASS: Component 2 — All 19 cells C2:C20 are right-aligned (0.4 pts)")
            total_score += 0.4
        elif right_count > 0:
            print(f"PARTIAL: Component 2 — Only {right_count}/19 cells in C2:C20 are right-aligned. Failures: {align_failures[:5]}")
        else:
            print(f"FAIL: Component 2 — No cells in C2:C20 are right-aligned. Sample: {align_failures[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: C1 (header 'Price') is center-aligned (0.2 points)
    # This FAILS on initial (C1 is 'left') and PASSES on golden (C1 is 'center')
    try:
        c1 = ws['C1']
        c1_align = c1.alignment.horizontal
        if c1_align == 'center':
            print(f"PASS: Component 3 — C1 is center-aligned (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — C1 alignment is {repr(c1_align)}, expected 'center'")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
