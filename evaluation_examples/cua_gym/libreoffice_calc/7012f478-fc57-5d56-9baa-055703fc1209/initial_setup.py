"""
Initial Setup: Format cells C2:C20 as currency and align headers/data
Task ID: calc_fmt_number_and_align_combined_064
Domain: libreoffice_calc

Creates a Price List spreadsheet with items and prices.
C2:C20 use 'General' format and are left-aligned (not yet formatted).
C1 is left-aligned (not yet centered).
"""

import openpyxl
from openpyxl.styles import Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_number_and_align_combined_064'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Price List'

    # Headers in row 1 - left-aligned (default, not yet centered)
    headers = ['Item', 'SKU', 'Price']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.alignment = Alignment(horizontal='left')

    # Realistic product data - 19 rows (rows 2-20)
    data = [
        ('Wireless Bluetooth Headphones', 'SKU-10021', 29.99),
        ('Laptop Stand Aluminum', 'SKU-10022', 149.5),
        ('USB-C Hub 7-Port', 'SKU-10023', 5.0),
        ('Mechanical Keyboard TKL', 'SKU-10024', 89.95),
        ('Ergonomic Mouse Wireless', 'SKU-10025', 45.0),
        ('Monitor LED 27-inch', 'SKU-10026', 319.99),
        ('Webcam HD 1080p', 'SKU-10027', 62.5),
        ('Desk Lamp LED Adjustable', 'SKU-10028', 38.75),
        ('Cable Organizer Bundle', 'SKU-10029', 12.0),
        ('HDMI Cable 6ft', 'SKU-10030', 8.49),
        ('SSD External 1TB', 'SKU-10031', 109.99),
        ('Portable Charger 20000mAh', 'SKU-10032', 54.0),
        ('Screen Cleaning Kit', 'SKU-10033', 9.95),
        ('USB Flash Drive 256GB', 'SKU-10034', 21.0),
        ('Smart Surge Protector', 'SKU-10035', 47.5),
        ('Noise Cancelling Earbuds', 'SKU-10036', 134.0),
        ('Laptop Sleeve 15-inch', 'SKU-10037', 24.99),
        ('Wireless Charging Pad', 'SKU-10038', 33.0),
        ('Gaming Controller USB', 'SKU-10039', 59.95),
    ]

    for r, (item, sku, price) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=item)
        ws.cell(row=r, column=2, value=sku)
        # Price cell: General format, left-aligned (not yet formatted)
        price_cell = ws.cell(row=r, column=3, value=price)
        price_cell.number_format = 'General'
        price_cell.alignment = Alignment(horizontal='left')

    # Column widths for readability
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Price List')
    print(f'  Rows: 1 header + 19 data rows (rows 2-20)')
    print(f'  C2:C20: General format, left-aligned')
    print(f'  C1: left-aligned (not centered)')


create_initial()
