"""
Initial Setup: INDIRECT function task - create spreadsheet with sheet names
              and cell addresses in Summary, with data sheets Jan/Feb/Mar
Task ID: calc_fma_indirect_008
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_fma_indirect_008'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Summary ----
    ws_summary = wb.active
    ws_summary.title = 'Summary'

    # Headers - row 1
    header_font = Font(bold=True)
    ws_summary['A1'] = 'Sheet Name'
    ws_summary['A1'].font = header_font
    ws_summary['B1'] = 'Cell Address'
    ws_summary['B1'].font = header_font
    ws_summary['C1'] = 'Value'
    ws_summary['C1'].font = header_font

    # Data rows 2-7: sheet names and cell addresses
    lookup_data = [
        ('Jan', 'B5'),
        ('Feb', 'B5'),
        ('Mar', 'B5'),
        ('Jan', 'C7'),
        ('Feb', 'C7'),
        ('Mar', 'C7'),
    ]
    for i, (sheet_name, cell_addr) in enumerate(lookup_data, start=2):
        ws_summary.cell(row=i, column=1, value=sheet_name)
        ws_summary.cell(row=i, column=2, value=cell_addr)
        # Column C (Value) intentionally left empty — task will fill with INDIRECT formulas

    # Column widths for readability
    ws_summary.column_dimensions['A'].width = 14
    ws_summary.column_dimensions['B'].width = 14
    ws_summary.column_dimensions['C'].width = 14

    # ---- Sheet 2: Jan ----
    # Layout: A=Product, B=Online Sales, C=In-Store Sales, D=Total Units, E=Net Profit
    # Requirement: B5=42500, C7=38900
    ws_jan = wb.create_sheet('Jan')

    jan_headers = ['Product', 'Online Sales', 'In-Store Sales', 'Total Units', 'Net Profit']
    for col, h in enumerate(jan_headers, 1):
        cell = ws_jan.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # row 2..11 (10 rows of data)
    # row 5: B5=42500 (Online Sales for row 5)
    # row 7: C7=38900 (In-Store Sales for row 7)
    jan_data = [
        ['Alpha Widget',    38200, 35100, 152, 19800],  # row 2
        ['Beta Gadget',     29500, 27400, 118, 15300],  # row 3
        ['Gamma Device',    41300, 37600, 165, 21600],  # row 4
        ['Delta Module',    42500, 40200, 170, 22400],  # row 5 — B5=42500
        ['Epsilon Tool',    35700, 33800, 143, 18700],  # row 6
        ['Zeta Component',  27600, 38900, 110, 14400],  # row 7 — C7=38900
        ['Eta System',      34100, 31500, 136, 17900],  # row 8
        ['Theta Unit',      31400, 29200, 126, 16300],  # row 9
        ['Iota Part',       44600, 42100, 178, 23400],  # row 10
        ['Kappa Assembly',  33100, 30800, 132, 17300],  # row 11
    ]
    for r, row_data in enumerate(jan_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_jan.cell(row=r, column=c, value=val)

    for col_letter in ['A', 'B', 'C', 'D', 'E']:
        ws_jan.column_dimensions[col_letter].width = 16

    # ---- Sheet 3: Feb ----
    # Requirement: B5=44100, C7=41200
    ws_feb = wb.create_sheet('Feb')

    feb_headers = ['Product', 'Online Sales', 'In-Store Sales', 'Total Units', 'Net Profit']
    for col, h in enumerate(feb_headers, 1):
        cell = ws_feb.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    feb_data = [
        ['Alpha Widget',    39800, 37100, 159, 20700],  # row 2
        ['Beta Gadget',     30200, 28600, 121, 15600],  # row 3
        ['Gamma Device',    43100, 39800, 172, 22600],  # row 4
        ['Delta Module',    44100, 41600, 176, 23200],  # row 5 — B5=44100
        ['Epsilon Tool',    36400, 34700, 146, 19000],  # row 6
        ['Zeta Component',  28300, 41200, 113, 14800],  # row 7 — C7=41200
        ['Eta System',      35600, 33100, 142, 18500],  # row 8
        ['Theta Unit',      32100, 30000, 128, 16700],  # row 9
        ['Iota Part',       46200, 43500, 185, 24200],  # row 10
        ['Kappa Assembly',  34700, 32400, 139, 18100],  # row 11
    ]
    for r, row_data in enumerate(feb_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_feb.cell(row=r, column=c, value=val)

    for col_letter in ['A', 'B', 'C', 'D', 'E']:
        ws_feb.column_dimensions[col_letter].width = 16

    # ---- Sheet 4: Mar ----
    # Requirement: B5=46800, C7=43500
    ws_mar = wb.create_sheet('Mar')

    mar_headers = ['Product', 'Online Sales', 'In-Store Sales', 'Total Units', 'Net Profit']
    for col, h in enumerate(mar_headers, 1):
        cell = ws_mar.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    mar_data = [
        ['Alpha Widget',    41500, 38800, 166, 21700],  # row 2
        ['Beta Gadget',     32400, 30500, 130, 16800],  # row 3
        ['Gamma Device',    45200, 41700, 181, 23700],  # row 4
        ['Delta Module',    46800, 44200, 187, 24600],  # row 5 — B5=46800
        ['Epsilon Tool',    38200, 36300, 153, 20000],  # row 6
        ['Zeta Component',  30100, 43500, 120, 15700],  # row 7 — C7=43500
        ['Eta System',      37400, 34900, 149, 19600],  # row 8
        ['Theta Unit',      34200, 31800, 137, 17800],  # row 9
        ['Iota Part',       48700, 45900, 195, 25500],  # row 10
        ['Kappa Assembly',  36300, 33900, 145, 19000],  # row 11
    ]
    for r, row_data in enumerate(mar_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_mar.cell(row=r, column=c, value=val)

    for col_letter in ['A', 'B', 'C', 'D', 'E']:
        ws_mar.column_dimensions[col_letter].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: Summary, Jan, Feb, Mar')
    print('Key values: Jan B5=42500, Jan C7=38900')
    print('            Feb B5=44100, Feb C7=41200')
    print('            Mar B5=46800, Mar C7=43500')
    print('Summary C2:C7 intentionally empty (task target)')


create_initial()
