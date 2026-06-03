"""
Reward Script: Use INDIRECT function to retrieve values from dynamic cell references
Task ID: calc_fma_indirect_008
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): INDIRECT formulas present in C2:C4 (Jan/Feb/Mar B5 references)
  Component 2 (0.3): INDIRECT formulas present in C5:C7 (Jan/Feb/Mar C7 references)
  Component 3 (0.2): All 6 formulas use correct INDIRECT pattern with dynamic A&"."&B references
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fma_indirect_008'


def is_indirect_formula(val):
    """Check if value is an INDIRECT formula string."""
    if not isinstance(val, str):
        return False
    return val.strip().upper().startswith('=INDIRECT(')


def has_dynamic_ref_pattern(val):
    """Check if the INDIRECT formula uses the A&'.'&B dynamic reference pattern."""
    if not isinstance(val, str):
        return False
    # Match pattern: =INDIRECT(A{n}&"."&B{n}) with possible variations in spacing/case
    # Should reference both column A and column B concatenated with "."
    v = val.strip().upper()
    has_indirect = v.startswith('=INDIRECT(')
    has_concat = '&' in v
    has_dot_sep = '"."' in v or "'.'".upper() in v
    return has_indirect and has_concat and has_dot_sep


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    The task requires adding INDIRECT formulas in Summary!C2:C7 to retrieve
    values from cross-sheet cell references specified in columns A and B.

    Initial state: C2:C7 are all empty (None)
    Golden state: C2:C7 contain =INDIRECT(A{n}&"."&B{n}) formulas
    """
    total_score = 0.0

    # Precondition gate: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: Summary sheet must exist
    if 'Summary' not in wb.sheetnames:
        print("CRITICAL: 'Summary' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Summary']

    # Component 1: INDIRECT formulas in C2:C4 (rows 2-4, Jan/Feb/Mar B5 references)
    # In initial file: C2:C4 are None → FAILS
    # In golden file:  C2:C4 have =INDIRECT(A{n}&"."&B{n}) formulas → PASSES
    try:
        c2_val = ws.cell(row=2, column=3).value
        c3_val = ws.cell(row=3, column=3).value
        c4_val = ws.cell(row=4, column=3).value

        c2_ok = is_indirect_formula(c2_val)
        c3_ok = is_indirect_formula(c3_val)
        c4_ok = is_indirect_formula(c4_val)

        count_c2_c4 = sum([c2_ok, c3_ok, c4_ok])
        comp1_score = round(count_c2_c4 / 3 * 0.5, 4)

        if count_c2_c4 == 3:
            print(f"PASS: Component 1 — All 3 INDIRECT formulas present in C2:C4 (0.5 pts)")
            print(f"  C2={repr(c2_val)}, C3={repr(c3_val)}, C4={repr(c4_val)}")
        elif count_c2_c4 > 0:
            print(f"PARTIAL: Component 1 — {count_c2_c4}/3 INDIRECT formulas in C2:C4 ({comp1_score} pts)")
            print(f"  C2={repr(c2_val)}, C3={repr(c3_val)}, C4={repr(c4_val)}")
        else:
            print(f"FAIL: Component 1 — No INDIRECT formulas in C2:C4")
            print(f"  C2={repr(c2_val)}, C3={repr(c3_val)}, C4={repr(c4_val)}")
            comp1_score = 0.0

        if comp1_score > 0:
            total_score += comp1_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: INDIRECT formulas in C5:C7 (rows 5-7, Jan/Feb/Mar C7 references)
    # In initial file: C5:C7 are None → FAILS
    # In golden file:  C5:C7 have =INDIRECT(A{n}&"."&B{n}) formulas → PASSES
    try:
        c5_val = ws.cell(row=5, column=3).value
        c6_val = ws.cell(row=6, column=3).value
        c7_val = ws.cell(row=7, column=3).value

        c5_ok = is_indirect_formula(c5_val)
        c6_ok = is_indirect_formula(c6_val)
        c7_ok = is_indirect_formula(c7_val)

        count_c5_c7 = sum([c5_ok, c6_ok, c7_ok])
        comp2_score = round(count_c5_c7 / 3 * 0.3, 4)

        if count_c5_c7 == 3:
            print(f"PASS: Component 2 — All 3 INDIRECT formulas present in C5:C7 (0.3 pts)")
            print(f"  C5={repr(c5_val)}, C6={repr(c6_val)}, C7={repr(c7_val)}")
        elif count_c5_c7 > 0:
            print(f"PARTIAL: Component 2 — {count_c5_c7}/3 INDIRECT formulas in C5:C7 ({comp2_score} pts)")
            print(f"  C5={repr(c5_val)}, C6={repr(c6_val)}, C7={repr(c7_val)}")
        else:
            print(f"FAIL: Component 2 — No INDIRECT formulas in C5:C7")
            print(f"  C5={repr(c5_val)}, C6={repr(c6_val)}, C7={repr(c7_val)}")
            comp2_score = 0.0

        if comp2_score > 0:
            total_score += comp2_score
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: All 6 formulas use correct INDIRECT dynamic reference pattern
    # (A{n}&"."&B{n} pattern - dynamic cross-sheet reference using "." separator)
    # Only awards points when all 6 formulas are present AND use correct pattern
    # In initial file: No formulas at all → FAILS
    # In golden file:  All 6 formulas present with correct pattern → PASSES
    try:
        all_vals = [ws.cell(row=r, column=3).value for r in range(2, 8)]
        formulas_present = [v for v in all_vals if is_indirect_formula(v)]
        dynamic_pattern_correct = [v for v in all_vals if has_dynamic_ref_pattern(v)]

        all_formulas_with_correct_pattern = (
            len(formulas_present) == 6 and len(dynamic_pattern_correct) == 6
        )

        if all_formulas_with_correct_pattern:
            print(f"PASS: Component 3 — All 6 formulas use correct INDIRECT(A&'.'&B) pattern (0.2 pts)")
            total_score += 0.2
        elif len(formulas_present) == 6 and len(dynamic_pattern_correct) < 6:
            print(f"FAIL: Component 3 — 6 INDIRECT formulas present but {len(dynamic_pattern_correct)}/6 "
                  f"use correct A&'.'&B dynamic reference pattern")
            print(f"  Formulas: {all_vals}")
        elif len(formulas_present) > 0:
            print(f"FAIL: Component 3 — Only {len(formulas_present)}/6 INDIRECT formulas present; "
                  f"{len(dynamic_pattern_correct)}/6 use correct pattern")
        else:
            print(f"FAIL: Component 3 — No INDIRECT formulas found in C2:C7")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
