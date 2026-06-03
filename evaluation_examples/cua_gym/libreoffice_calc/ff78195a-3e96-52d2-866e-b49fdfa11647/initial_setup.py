"""
Initial Setup: Create time tracking spreadsheet with 480 rows of data
Task ID: calc_pivot_092
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_092'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'TimeTracking'

    # --- Headers ---
    headers = ['EntryID', 'Date', 'TeamMember', 'Project', 'TotalHours', 'BillableHours']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Data Setup ---
    team_members = [
        'Sarah Chen',
        'Marcus Johnson',
        'Priya Patel',
        'James O\'Brien',
        'Aisha Williams',
        'Carlos Rivera',
    ]

    projects = [
        'Phoenix Migration',
        'Atlas Dashboard',
        'Nebula API',
        'Orion Analytics',
        'Titan Infrastructure',
        'Horizon Mobile',
    ]

    months = [
        (2024, 1), (2024, 2), (2024, 3),
        (2024, 4), (2024, 5), (2024, 6),
    ]

    # We need 480 rows total, grand total TotalHours = 3840, overall billable rate ~0.75
    # 480 rows => average TotalHours per row = 8.0
    # We'll generate 80 entries per member (480 / 6 = 80)
    # Each member gets ~80 entries spread across 6 months (~13-14 per month)

    # Build entries deterministically to hit targets
    rows_data = []
    entry_id = 1

    for member_idx, member in enumerate(team_members):
        # Each member: 80 entries, total hours = 640 (so grand total = 640*6 = 3840)
        # Billable = ~75% of total = ~480 per member
        member_total = 0
        member_billable = 0
        target_total = 640
        target_billable = 480  # 75%

        entries_per_month = [13, 13, 14, 13, 13, 14]  # sums to 80

        for month_idx, (year, month) in enumerate(months):
            n_entries = entries_per_month[month_idx]
            for j in range(n_entries):
                day = random.randint(1, 28)
                date_str = f'{month:02d}/{day:02d}/{year}'

                project = projects[(member_idx + j) % len(projects)]

                # Distribute hours to hit targets
                remaining_entries = 80 - len([r for r in rows_data if r[2] == member])
                if remaining_entries <= 1:
                    total_h = target_total - member_total
                else:
                    avg_remaining = (target_total - member_total) / remaining_entries
                    total_h = round(random.gauss(avg_remaining, 1.5), 1)
                    total_h = max(4.0, min(12.0, total_h))
                    total_h = round(total_h, 1)

                member_total += total_h

                # Billable hours: ~75% of total
                remaining_entries_b = 80 - len([r for r in rows_data if r[2] == member])
                if remaining_entries_b <= 1:
                    bill_h = target_billable - member_billable
                else:
                    bill_h = round(total_h * random.uniform(0.6, 0.9), 1)
                    bill_h = min(bill_h, total_h)

                member_billable += bill_h

                rows_data.append([
                    entry_id, date_str, member, project,
                    round(total_h, 1), round(bill_h, 1),
                ])
                entry_id += 1

    # Write data rows
    for r, row_data in enumerate(rows_data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 16

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Total rows: {len(rows_data)}')

    # Compute some stats for verification
    grand_total = sum(r[4] for r in rows_data)
    grand_billable = sum(r[5] for r in rows_data)
    print(f'Grand TotalHours: {grand_total}')
    print(f'Grand BillableHours: {grand_billable}')
    print(f'Overall Billable Rate: {grand_billable/grand_total:.4f}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
