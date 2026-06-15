"""
Reward Script: Create pivot table from project time tracking data
Task ID: calc_pivot_092
Domain: libreoffice_calc
Scoring:
  Component 1 (0.15): PivotTable sheet exists
  Component 2 (0.25): TotalHours section structure (headers, team members, months)
  Component 3 (0.25): TotalHours Grand Total == 3840
  Component 4 (0.15): BillableRate section exists with correct structure
  Component 5 (0.20): BillableRate overall value ~0.75
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_092'

EXPECTED_MEMBERS = [
    'Sarah Chen', 'Marcus Johnson', 'Priya Patel',
    "James O'Brien", 'Aisha Williams', 'Carlos Rivera'
]
EXPECTED_MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun']


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: PivotTable sheet exists (0.15 points)
    # This is the fundamental task-introduced change -- initial has only TimeTracking
    try:
        sheet_names = wb.sheetnames
        pivot_sheet = None
        for sn in sheet_names:
            if 'pivot' in sn.lower():
                pivot_sheet = sn
                break
        if pivot_sheet:
            ws = wb[pivot_sheet]
            # Verify it has meaningful content (not just an empty sheet)
            if ws.max_row >= 10 and ws.max_column >= 3:
                print(f"PASS: Component 1 -- PivotTable sheet '{pivot_sheet}' exists with content (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 1 -- Sheet '{pivot_sheet}' exists but has insufficient content (max_row={ws.max_row}, max_col={ws.max_column})")
        else:
            print(f"FAIL: Component 1 -- No sheet with 'pivot' in name found. Sheets: {sheet_names}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    if pivot_sheet is None:
        print("\nScore: 0.0/1.0")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[pivot_sheet]

    # Component 2: TotalHours section structure (0.25 points)
    # Check that team members appear as row labels and months as column headers
    try:
        # Scan for month headers in the first 15 rows
        month_header_row = None
        for r in range(1, min(16, ws.max_row + 1)):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            # Check if this row contains month abbreviations
            str_vals = [str(v).strip() if v else '' for v in row_vals]
            month_matches = sum(1 for m in EXPECTED_MONTHS if m in str_vals)
            if month_matches >= 4:  # At least 4 of 6 months found
                month_header_row = r
                break

        if month_header_row is None:
            print("FAIL: Component 2 -- Could not find month headers (Jan-Jun) in first 15 rows")
        else:
            # Check team members appear as row labels below the header
            member_col = None
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=month_header_row, column=c).value
                if val and ('member' in str(val).lower() or 'team' in str(val).lower()):
                    member_col = c
                    break
            if member_col is None:
                member_col = 1  # Default to column A

            found_members = []
            for r in range(month_header_row + 1, min(month_header_row + 15, ws.max_row + 1)):
                val = ws.cell(row=r, column=member_col).value
                if val and str(val).strip() in EXPECTED_MEMBERS:
                    found_members.append(str(val).strip())

            if len(found_members) >= 5:  # At least 5 of 6 members
                print(f"PASS: Component 2 -- TotalHours structure found: {len(found_members)} members, months at row {month_header_row} (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 2 -- Only found {len(found_members)} team members (need >= 5): {found_members}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: TotalHours Grand Total == 3840 (0.25 points)
    # The task context specifies Grand Total TotalHours = 3840
    try:
        grand_total_found = False
        # Search for a cell containing 3840 (or close to it) in the pivot area
        for r in range(1, min(25, ws.max_row + 1)):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    try:
                        num_val = float(val)
                        if abs(num_val - 3840) < 1.0:
                            # Verify the row or column label says "Grand Total" or similar
                            row_label = ws.cell(row=r, column=1).value
                            col_label = ws.cell(row=1, column=c).value if c > 1 else None
                            header_row_label = None
                            if month_header_row:
                                header_row_label = ws.cell(row=month_header_row, column=c).value
                            is_grand = False
                            if row_label and 'grand' in str(row_label).lower():
                                is_grand = True
                            if header_row_label and 'grand' in str(header_row_label).lower():
                                is_grand = True
                            if header_row_label and 'total' in str(header_row_label).lower():
                                is_grand = True
                            # Also accept if it's in the last column of the header area
                            if c == ws.max_column:
                                is_grand = True

                            if is_grand:
                                grand_total_found = True
                                print(f"PASS: Component 3 -- Grand Total TotalHours = {num_val} at ({r},{c}) (0.25 pts)")
                                total_score += 0.25
                                break
                    except (ValueError, TypeError):
                        pass
            if grand_total_found:
                break

        if not grand_total_found:
            print("FAIL: Component 3 -- Grand Total of 3840 not found in pivot table area")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: BillableRate section exists (0.15 points)
    # There should be a second section with BillableRate/billable data
    try:
        billable_section_found = False
        billable_start_row = None
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if val and isinstance(val, str):
                    val_lower = val.lower()
                    if 'billable' in val_lower and 'rate' in val_lower:
                        billable_section_found = True
                        billable_start_row = r
                        break
            if billable_section_found:
                break

        if billable_section_found:
            # Verify it has data below (team members with decimal values)
            data_rows_found = 0
            for r in range(billable_start_row + 1, min(billable_start_row + 15, ws.max_row + 1)):
                row_label = ws.cell(row=r, column=1).value
                if row_label and str(row_label).strip() in EXPECTED_MEMBERS:
                    # Check that there are decimal values (rates between 0 and 1)
                    for c in range(2, ws.max_column + 1):
                        v = ws.cell(row=r, column=c).value
                        if v is not None:
                            try:
                                fv = float(v)
                                if 0 < fv < 1.0:
                                    data_rows_found += 1
                                    break
                            except (ValueError, TypeError):
                                pass

            if data_rows_found >= 4:
                print(f"PASS: Component 4 -- BillableRate section found at row {billable_start_row} with {data_rows_found} member rows (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 4 -- BillableRate header found but only {data_rows_found} data rows with rate values")
        else:
            print("FAIL: Component 4 -- No BillableRate section header found in pivot table")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # Component 5: BillableRate overall value ~0.75 (0.20 points)
    # The task context says expected overall billable rate around 0.75
    try:
        overall_rate_found = False
        # Search for Grand Total row in the billable section
        if billable_start_row:
            for r in range(billable_start_row + 1, min(billable_start_row + 15, ws.max_row + 1)):
                row_label = ws.cell(row=r, column=1).value
                if row_label and 'grand' in str(row_label).lower():
                    # Check the last column for overall rate
                    for c in range(ws.max_column, 1, -1):
                        val = ws.cell(row=r, column=c).value
                        if val is not None:
                            try:
                                rate = float(val)
                                if abs(rate - 0.75) < 0.05:
                                    overall_rate_found = True
                                    print(f"PASS: Component 5 -- Overall BillableRate = {rate} (expected ~0.75) (0.20 pts)")
                                    total_score += 0.20
                                else:
                                    print(f"FAIL: Component 5 -- Overall BillableRate = {rate}, expected ~0.75 (tolerance 0.05)")
                                break
                            except (ValueError, TypeError):
                                pass
                    break

        if not overall_rate_found and total_score < 0.75:
            # Also scan broadly for 0.75 in the lower part of the sheet
            for r in range(12, ws.max_row + 1):
                row_label = ws.cell(row=r, column=1).value
                if row_label and ('grand' in str(row_label).lower() or 'total' in str(row_label).lower()):
                    for c in range(ws.max_column, 1, -1):
                        val = ws.cell(row=r, column=c).value
                        if val is not None:
                            try:
                                rate = float(val)
                                if abs(rate - 0.75) < 0.05:
                                    overall_rate_found = True
                                    print(f"PASS: Component 5 -- Overall BillableRate = {rate} found at ({r},{c}) (0.20 pts)")
                                    total_score += 0.20
                                    break
                            except (ValueError, TypeError):
                                pass
                    if overall_rate_found:
                        break

        if not overall_rate_found:
            print("FAIL: Component 5 -- Could not find overall BillableRate ~0.75 in Grand Total row")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
