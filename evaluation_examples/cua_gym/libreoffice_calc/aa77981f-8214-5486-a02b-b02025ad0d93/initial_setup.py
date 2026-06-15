"""
Initial Setup: Project Risk Register
Task ID: calc_ops_project_risk_register_061
Domain: libreoffice_calc

Creates a risk register spreadsheet with 25 risk entries.
Columns F (Risk Score), G (Risk Level), and C/J dropdowns are empty —
agent must add formulas, dropdowns, conditional formatting, and sorting.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_project_risk_register_061'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'RiskRegister'

    # --- Headers ---
    headers = [
        'Risk ID', 'Risk Description', 'Category',
        'Probability 1-5', 'Impact 1-5', 'Risk Score',
        'Risk Level', 'Owner', 'Mitigation', 'Status'
    ]
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF')
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    ws.row_dimensions[1].height = 28

    # --- Column widths ---
    col_widths = [10, 45, 14, 16, 12, 12, 12, 20, 45, 14]
    for i, w in enumerate(col_widths, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- 25 Risk entries ---
    # Columns: Risk ID, Description, Category(empty), Prob, Impact, Score(empty), Level(empty), Owner, Mitigation, Status(empty)
    risk_data = [
        ('R-001', 'Key vendor fails to deliver critical hardware components on schedule',
         None, 4, 5, None, None,
         'James Whitfield',
         'Identify backup vendors; maintain 4-week buffer stock for critical components',
         None),
        ('R-002', 'Scope creep due to undocumented stakeholder requirements',
         None, 5, 4, None, None,
         'Priya Sharma',
         'Formal change control process; weekly stakeholder sign-off on scope',
         None),
        ('R-003', 'Lead developer resigns during critical sprint',
         None, 3, 5, None, None,
         'Chen Wei',
         'Knowledge transfer sessions; pair programming and documentation standards',
         None),
        ('R-004', 'Regulatory approval delayed due to compliance gaps',
         None, 3, 4, None, None,
         'Amara Osei',
         'Engage compliance consultant; conduct internal audit 6 weeks before submission',
         None),
        ('R-005', 'Cloud infrastructure outage affecting production environment',
         None, 2, 5, None, None,
         'Tobias Müller',
         'Multi-region failover configured; RTO target 15 minutes; monthly DR drills',
         None),
        ('R-006', 'Budget overrun due to underestimated third-party licensing costs',
         None, 4, 3, None, None,
         'Linda Nakamura',
         'Procurement review at project initiation; contingency fund 15% of license budget',
         None),
        ('R-007', 'Data migration errors causing corrupted customer records',
         None, 3, 5, None, None,
         'Samuel Adeyemi',
         'Staged migration with reconciliation checkpoints; rollback plan prepared',
         None),
        ('R-008', 'Security vulnerability discovered in core authentication module',
         None, 2, 5, None, None,
         'Fatima Al-Hassan',
         'Penetration testing before go-live; immediate patch protocol established',
         None),
        ('R-009', 'Integration failure between legacy ERP and new payment gateway',
         None, 4, 4, None, None,
         'Robert Andersen',
         'Dedicated integration testing environment; API contract testing automated',
         None),
        ('R-010', 'Project sponsor changes priorities mid-project',
         None, 3, 4, None, None,
         'Yuki Tanaka',
         'Steering committee charter with defined escalation path; monthly sponsor reviews',
         None),
        ('R-011', 'Loss of critical project documentation due to system failure',
         None, 2, 3, None, None,
         'Diana Petrov',
         'Daily automated backups to separate cloud storage; version control enforced',
         None),
        ('R-012', 'Team members unfamiliar with new technology stack',
         None, 4, 3, None, None,
         'Carlos Mendes',
         'Training budget allocated; senior contractor hired for knowledge transfer',
         None),
        ('R-013', 'Third-party API deprecated without adequate notice',
         None, 2, 4, None, None,
         'Ngozi Eze',
         'API versioning monitored; abstraction layer implemented to ease migration',
         None),
        ('R-014', 'Delays in obtaining user acceptance testing sign-off',
         None, 3, 3, None, None,
         'Erik Lindqvist',
         'UAT schedule locked 8 weeks prior; dedicated UAT environment provisioned',
         None),
        ('R-015', 'Currency exchange rate fluctuations affecting overseas contractor costs',
         None, 3, 3, None, None,
         'Mei-Ling Zhou',
         'Contracts denominated in local currency; hedging options reviewed quarterly',
         None),
        ('R-016', 'Intellectual property dispute over reused open-source components',
         None, 2, 4, None, None,
         'Kwame Asante',
         'Legal review of open-source licenses; FOSS inventory maintained',
         None),
        ('R-017', 'Performance testing reveals system cannot handle peak load',
         None, 3, 4, None, None,
         'Ingrid Svensson',
         'Load testing scheduled 10 weeks before launch; auto-scaling policies configured',
         None),
        ('R-018', 'Miscommunication between onshore and offshore teams causing rework',
         None, 4, 2, None, None,
         'Arjun Patel',
         'Daily stand-up overlap window; shared collaboration tools; clear RACI matrix',
         None),
        ('R-019', 'Environmental regulation changes impacting product requirements',
         None, 2, 3, None, None,
         'Sofía Rodríguez',
         'Legal counsel monitors regulatory landscape; modular product design for adaptability',
         None),
        ('R-020', 'Unexpected increase in raw material prices',
         None, 3, 3, None, None,
         'Hassan Ibrahim',
         'Long-term supplier contracts signed; substitute materials evaluated',
         None),
        ('R-021', 'User interface redesign rejected by end users in testing',
         None, 2, 3, None, None,
         'Anika Hoffmann',
         'Early prototypes shared with user focus groups; iterative design reviews',
         None),
        ('R-022', 'Network connectivity issues at remote client site',
         None, 2, 2, None, None,
         'Patrick O\'Brien',
         'Offline mode designed into application; VPN and SD-WAN options evaluated',
         None),
        ('R-023', 'Competitor launches similar product before our go-live date',
         None, 3, 2, None, None,
         'Natalie Kim',
         'Accelerated MVP scope defined; differentiation features prioritised in roadmap',
         None),
        ('R-024', 'Insufficient internal testing resources during final QA phase',
         None, 3, 2, None, None,
         'Luca Ferrari',
         'Test automation coverage target 80%; contractors available for surge capacity',
         None),
        ('R-025', 'Organisational restructuring disrupts project governance structure',
         None, 2, 2, None, None,
         'Miriam Goldstein',
         'Project charter includes continuity provisions; executive sponsor commitment documented',
         None),
    ]

    for r, row in enumerate(risk_data, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            if c in (1,):
                cell.alignment = Alignment(horizontal='center')
            if c in (4, 5):
                cell.alignment = Alignment(horizontal='center')

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('  Sheet: RiskRegister')
    print('  Rows: 1 header + 25 data rows')
    print('  Columns F, G empty (no formulas yet)')
    print('  Columns C, J empty (no dropdowns yet)')
    print('  No conditional formatting')


create_initial()
