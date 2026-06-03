"""
Reward Script: Build a project risk register with probability/impact scoring, risk classification, data validation dropdowns, and conditional formatting.
Task ID: calc_ops_project_risk_register_061
Domain: libreoffice_calc
Scoring:
  - Component 1: Risk Score formulas F2:F26 = D*E (0.25 pts)
  - Component 2: Risk Level formulas G2:G26 = IF classification (0.25 pts)
  - Component 3: Data validation dropdowns on C2:C26 (Category) and J2:J26 (Status) (0.25 pts)
  - Component 4: Conditional formatting on G column (High/Medium/Low colors) and entire row (0.25 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_project_risk_register_061'

def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify RiskRegister sheet exists (precondition gate — not scored)
    if 'RiskRegister' not in wb.sheetnames:
        print("CRITICAL: Sheet 'RiskRegister' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['RiskRegister']

    # Component 1: Risk Score formulas in F2:F26 (=D*E pattern) (0.25 pts)
    # Initial: F2:F26 are all empty (None)
    # Golden: F2:F26 must each contain =D{row}*E{row} formula
    try:
        formula_count = 0
        total_rows = 25  # rows 2-26
        for row in range(2, 27):
            f_val = ws.cell(row=row, column=6).value
            if f_val is not None:
                f_str = str(f_val).strip().upper().replace(' ', '')
                # Accept =D2*E2 or =E2*D2 form
                if f_str in (f'=D{row}*E{row}', f'=E{row}*D{row}'):
                    formula_count += 1

        if formula_count == total_rows:
            print(f"PASS: Component 1 — All {total_rows} Risk Score formulas (F2:F26) are =D*E (0.25 pts)")
            total_score += 0.25
        elif formula_count >= 20:
            print(f"PARTIAL: Component 1 — {formula_count}/{total_rows} Risk Score formulas found (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 — Only {formula_count}/{total_rows} Risk Score formulas found in F2:F26")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Risk Level IF formulas in G2:G26 (0.25 pts)
    # Initial: G2:G26 are all empty (None)
    # Golden: G2:G26 must contain =IF(F>=15,"High",IF(F>=8,"Medium","Low")) pattern
    try:
        level_formula_count = 0
        total_rows = 25
        for row in range(2, 27):
            g_val = ws.cell(row=row, column=7).value
            if g_val is not None:
                g_str = str(g_val).strip().upper().replace(' ', '').replace('"', '').replace("'", '')
                # Must reference F column, use IF, thresholds 15 and 8, classify High/Medium/Low
                if (f'F{row}' in g_str and 'IF' in g_str and '15' in g_str
                        and '8' in g_str and 'HIGH' in g_str and 'MEDIUM' in g_str and 'LOW' in g_str):
                    level_formula_count += 1

        if level_formula_count == total_rows:
            print(f"PASS: Component 2 — All {total_rows} Risk Level formulas (G2:G26) are correct IF classifications (0.25 pts)")
            total_score += 0.25
        elif level_formula_count >= 20:
            print(f"PARTIAL: Component 2 — {level_formula_count}/{total_rows} Risk Level formulas found (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 — Only {level_formula_count}/{total_rows} Risk Level IF formulas found in G2:G26")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Data validation dropdowns on C2:C26 and J2:J26 (0.25 pts)
    # Initial: no data validations at all
    # Golden: must have list DV on C column (category) and J column (status)
    try:
        validations = ws.data_validations.dataValidation
        category_options = {'TECHNICAL', 'COMMERCIAL', 'OPERATIONAL', 'LEGAL', 'FINANCIAL', 'RESOURCE'}
        status_options = {'OPEN', 'MITIGATED', 'ACCEPTED', 'CLOSED', 'TRANSFERRED'}

        category_dv_count = sum(
            1 for dv in validations
            if dv.type == 'list' and dv.formula1
            and 'C' in str(dv.sqref).upper()
            and len(category_options & {v.strip() for v in dv.formula1.upper().replace('"', '').replace("'", '').split(',')}) >= 4
        )
        status_dv_count = sum(
            1 for dv in validations
            if dv.type == 'list' and dv.formula1
            and 'J' in str(dv.sqref).upper()
            and len(status_options & {v.strip() for v in dv.formula1.upper().replace('"', '').replace("'", '').split(',')}) >= 3
        )

        if category_dv_count >= 1:
            print(f"  Found category DV on column C")
        if status_dv_count >= 1:
            print(f"  Found status DV on column J")

        if category_dv_count >= 1 and status_dv_count >= 1:
            print(f"PASS: Component 3 — Both Category (C) and Status (J) data validation dropdowns present (0.25 pts)")
            total_score += 0.25
        elif category_dv_count >= 1 or status_dv_count >= 1:
            which = "Category (C)" if category_dv_count >= 1 else "Status (J)"
            print(f"PARTIAL: Component 3 — Only {which} dropdown found (0.12 pts)")
            total_score += 0.12
        else:
            print(f"FAIL: Component 3 — No data validation dropdowns found. Expected C2:C26 and J2:J26")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Conditional formatting on G column and entire row (0.25 pts)
    # Initial: no conditional formatting
    # Golden: must have CF rules on G column (High=red, Medium=orange, Low=green)
    #         and/or row-level CF (red row if $G='High')
    try:
        cf_rules = ws.conditional_formatting

        # Count CF rules on G column that reference High/Medium/Low
        g_col_cf_count = 0
        row_level_cf_count = 0

        for cf_range in cf_rules:
            cf_range_str = str(cf_range).upper()
            rules = cf_rules[cf_range]

            # G column CF: range contains 'G', formula references High/Medium/Low
            if 'G' in cf_range_str:
                for rule in rules:
                    formula_str = str(rule.formula).upper() if rule.formula else ''
                    if 'HIGH' in formula_str or 'MEDIUM' in formula_str or 'LOW' in formula_str:
                        g_col_cf_count += 1
                        print(f"  Found G column CF: range={cf_range}, formula={rule.formula}")

            # Row-level CF: range spans many columns (>=5) AND formula uses $G anchor with 'HIGH'
            col_count_in_range = sum(1 for col_letter in 'ABCDEFGHIJ' if col_letter in cf_range_str)
            if col_count_in_range >= 5:
                for rule in rules:
                    formula_str = str(rule.formula).upper() if rule.formula else ''
                    if 'HIGH' in formula_str and '$G' in formula_str:
                        row_level_cf_count += 1
                        print(f"  Found row-level CF: range={cf_range}, formula={rule.formula}")

        if g_col_cf_count >= 1 and row_level_cf_count >= 1:
            print(f"PASS: Component 4 — Both G column color CF ({g_col_cf_count} rules) and row-level High CF present (0.25 pts)")
            total_score += 0.25
        elif g_col_cf_count >= 1 or row_level_cf_count >= 1:
            which = "G column" if g_col_cf_count >= 1 else "row-level"
            print(f"PARTIAL: Component 4 — Only {which} conditional formatting found (0.12 pts)")
            total_score += 0.12
        else:
            print(f"FAIL: Component 4 — No conditional formatting found. Expected CF on G column and row level")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
