"""
Initial Setup: Define named ranges and build dashboard summary in LibreOffice Calc
Task ID: calc_ggf_035
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_035'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Data sheet ---
    ws = wb.active
    ws.title = 'Data'

    # Headers
    ws.cell(row=1, column=1, value='Month')
    ws.cell(row=1, column=2, value='Q1')
    ws.cell(row=1, column=3, value='Q2')
    ws.cell(row=1, column=4, value='Q3')
    ws.cell(row=1, column=5, value='Q4')

    # Dashboard headers in G1:H1
    ws.cell(row=1, column=7, value='Quarter')
    ws.cell(row=1, column=8, value='Total')

    # 13 rows of monthly data (rows 2-14)
    # Representing monthly revenue figures across four quarters
    monthly_data = [
        ['January',    12450, 15320, 18760, 22100],
        ['February',   11890, 14650, 17430, 21560],
        ['March',      13200, 16100, 19250, 23400],
        ['April',      12780, 15890, 18100, 22750],
        ['May',        14100, 16750, 20300, 24100],
        ['June',       13560, 15420, 19870, 23650],
        ['July',       12900, 16200, 18540, 22300],
        ['August',     13750, 15980, 19100, 23870],
        ['September',  14200, 16430, 20150, 24500],
        ['October',    13100, 15750, 18900, 22800],
        ['November',   12650, 16080, 19450, 23150],
        ['December',   14500, 17200, 21000, 25300],
        ['Adjustment', -1200,  -850,  -600,  -430],
    ]

    for r, row_data in enumerate(monthly_data, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        ws.cell(row=r, column=2, value=row_data[1])
        ws.cell(row=r, column=3, value=row_data[2])
        ws.cell(row=r, column=4, value=row_data[3])
        ws.cell(row=r, column=5, value=row_data[4])

    # G2:H6 intentionally left EMPTY — the task is to fill these with named-range formulas
    # No named ranges defined — the task is to create them

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
