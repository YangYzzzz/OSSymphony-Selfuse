"""
Reward Script: Define named ranges and build dashboard summary in LibreOffice Calc
Task ID: calc_ggf_035
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): Named ranges Q1-Q4 exist with correct references
  Component 2 (0.15): Named range FullYear exists with correct reference
  Component 3 (0.30): Dashboard G2:G5 labels + H2:H5 SUM formulas using named ranges
  Component 4 (0.20): Grand total cell uses =SUM(FullYear)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_035'


def persist_app_state(domain):
    """Attempt to save any unsaved LibreOffice state via Ctrl+S."""
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(1.0)
        print("PERSIST: ctrl+s sent for {}".format(domain))
    except Exception as e:
        print("PERSIST_WARN: save hook failed: {}".format(e))


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print("CRITICAL: Cannot load file {}: {}".format(file_path, e))
        print("REWARD: 0.0")
        return 0.0

    # Ensure 'Data' sheet exists
    if 'Data' not in wb.sheetnames:
        print("CRITICAL: 'Data' sheet not found. Sheets: {}".format(wb.sheetnames))
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Data']

    # Build a dict of defined names for easy lookup
    defined = {}
    for name, dn in wb.defined_names.items():
        dests = list(dn.destinations)
        if dests:
            sheet_title, coord = dests[0]
            defined[name.upper()] = (sheet_title, coord)

    # ---------------------------------------------------------------
    # Component 1: Named ranges Q1-Q4 exist with correct references (0.35 pts)
    # Each quarter is worth 0.35/4 = 0.0875 pts
    # ---------------------------------------------------------------
    try:
        expected_quarters = {
            'Q1': '$B$2:$B$14',
            'Q2': '$C$2:$C$14',
            'Q3': '$D$2:$D$14',
            'Q4': '$E$2:$E$14',
        }
        quarter_score = 0.0
        per_quarter = 0.35 / 4.0
        for qname, expected_range in expected_quarters.items():
            if qname.upper() in defined:
                sheet_title, coord = defined[qname.upper()]
                if coord == expected_range and sheet_title == 'Data':
                    print("PASS: Named range '{}' -> Data!{} (correct)".format(qname, coord))
                    quarter_score += per_quarter
                else:
                    print("FAIL: Named range '{}' -> {}!{}, expected Data!{}".format(
                        qname, sheet_title, coord, expected_range))
            else:
                print("FAIL: Named range '{}' not found".format(qname))
        total_score += quarter_score
        if quarter_score >= 0.34:
            print("PASS: Component 1 - All four quarter named ranges correct ({:.3f} pts)".format(quarter_score))
        else:
            print("PARTIAL: Component 1 - Quarter named ranges ({:.3f}/0.35 pts)".format(quarter_score))
    except Exception as e:
        print("ERROR: Component 1 - {}".format(e))

    # ---------------------------------------------------------------
    # Component 2: Named range FullYear exists with correct reference (0.15 pts)
    # ---------------------------------------------------------------
    try:
        if 'FULLYEAR' in defined:
            sheet_title, coord = defined['FULLYEAR']
            if coord == '$B$2:$E$14' and sheet_title == 'Data':
                print("PASS: Component 2 - FullYear -> Data!$B$2:$E$14 (0.15 pts)")
                total_score += 0.15
            else:
                print("FAIL: Component 2 - FullYear -> {}!{}, expected Data!$B$2:$E$14".format(
                    sheet_title, coord))
        else:
            print("FAIL: Component 2 - Named range 'FullYear' not found")
    except Exception as e:
        print("ERROR: Component 2 - {}".format(e))

    # ---------------------------------------------------------------
    # Component 3: Dashboard labels G2:G5 + SUM formulas H2:H5 using named ranges (0.30 pts)
    # Labels: 0.10 pts (0.025 each), Formulas: 0.20 pts (0.05 each)
    # ---------------------------------------------------------------
    try:
        quarter_labels = {2: 'Q1', 3: 'Q2', 4: 'Q3', 5: 'Q4'}
        label_score = 0.0
        formula_score = 0.0

        for row, expected_label in quarter_labels.items():
            # Check label in G column
            g_val = ws.cell(row=row, column=7).value  # G column
            if g_val is not None and str(g_val).strip().upper() == expected_label.upper():
                label_score += 0.025
                print("PASS: G{} = '{}' (correct label)".format(row, g_val))
            else:
                print("FAIL: G{} = '{}', expected '{}'".format(row, g_val, expected_label))

            # Check formula in H column - should use named range
            h_val = ws.cell(row=row, column=8).value  # H column
            if h_val is not None and isinstance(h_val, str):
                h_upper = h_val.upper().replace(' ', '')
                expected_formula = '=SUM({})'.format(expected_label).upper()
                if h_upper == expected_formula:
                    formula_score += 0.05
                    print("PASS: H{} = '{}' (correct named-range formula)".format(row, h_val))
                else:
                    print("FAIL: H{} = '{}', expected '=SUM({})'".format(row, h_val, expected_label))
            else:
                print("FAIL: H{} = '{}', expected a SUM formula using named range".format(row, h_val))

        total_score += label_score + formula_score
        comp3_total = label_score + formula_score
        if comp3_total >= 0.29:
            print("PASS: Component 3 - Dashboard labels and formulas ({:.3f} pts)".format(comp3_total))
        else:
            print("PARTIAL: Component 3 - Dashboard ({:.3f}/0.30 pts)".format(comp3_total))
    except Exception as e:
        print("ERROR: Component 3 - {}".format(e))

    # ---------------------------------------------------------------
    # Component 4: Grand total using =SUM(FullYear) (0.20 pts)
    # Look in the dashboard area (rows 6-8, columns G-H) for a grand total
    # ---------------------------------------------------------------
    try:
        grand_total_found = False
        # Search rows 6-10 in columns G and H for the grand total formula
        for row in range(6, 11):
            h_val = ws.cell(row=row, column=8).value  # H column
            if h_val is not None and isinstance(h_val, str):
                h_upper = h_val.upper().replace(' ', '')
                if h_upper == '=SUM(FULLYEAR)':
                    print("PASS: Component 4 - H{} = '{}' (grand total with FullYear) (0.20 pts)".format(row, h_val))
                    total_score += 0.20
                    grand_total_found = True
                    break
        if not grand_total_found:
            # Also check if it's in another column nearby
            for row in range(6, 11):
                for col in range(6, 10):  # F through I
                    val = ws.cell(row=row, column=col).value
                    if val is not None and isinstance(val, str):
                        v_upper = val.upper().replace(' ', '')
                        if v_upper == '=SUM(FULLYEAR)':
                            print("PASS: Component 4 - {}{} = '{}' (grand total with FullYear) (0.20 pts)".format(
                                chr(64 + col), row, val))
                            total_score += 0.20
                            grand_total_found = True
                            break
                if grand_total_found:
                    break
        if not grand_total_found:
            print("FAIL: Component 4 - No =SUM(FullYear) formula found in dashboard area")
    except Exception as e:
        print("ERROR: Component 4 - {}".format(e))

    final_score = min(round(total_score, 4), 1.0)
    print("\nScore: {}/1.0".format(final_score))
    print("REWARD: {}".format(final_score))
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = '{}/{}.xlsx'.format(WORKDIR, TASK_ID)
if not os.path.exists(file_path):
    print("File not found: {}".format(file_path))
    print("REWARD: 0.0")
else:
    verify_task(file_path)
