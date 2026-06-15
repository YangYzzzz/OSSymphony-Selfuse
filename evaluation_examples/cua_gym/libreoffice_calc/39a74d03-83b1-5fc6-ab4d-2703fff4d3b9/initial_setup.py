"""
Initial Setup: Billboard Hot 100 listening history task
Task ID: osworld_multi_apps_misc_014
Domain: libreoffice_calc + chrome (multi-app)

Creates my_songs.xlsx with listening history, opens LibreOffice Calc and Chrome.
Chrome is opened to billboard.com/charts/hot-100.
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_misc_014'
OUTPUT = f'{WORKDIR}/my_songs.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # Sheet: listening history
    ws = wb.active
    ws.title = 'my_songs'

    # Headers: Position, Title, Artist, Year
    headers = ['Position', 'Title', 'Artist', 'Year']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Listening history - a subset of well-known Billboard hits.
    # These are songs that the user has ALREADY listened to.
    # The agent must find top-30 pre-2024 songs NOT in this list.
    listened_songs = [
        # (Position_on_chart, Title, Artist, Year)
        # Songs the user HAS listened to (some from top 30, some not)
        (3,  'Flowers',                     'Miley Cyrus',              2023),
        (7,  'As It Was',                   'Harry Styles',             2022),
        (11, 'Blinding Lights',             'The Weeknd',               2020),
        (15, 'Shape of You',                'Ed Sheeran',               2017),
        (18, 'Uptown Funk',                 'Mark Ronson ft. Bruno Mars', 2015),
        (22, 'Rolling in the Deep',         'Adele',                    2011),
        (25, 'Happy',                       'Pharrell Williams',        2014),
        (28, 'Thinking Out Loud',           'Ed Sheeran',               2015),
        (32, 'Shake It Off',                'Taylor Swift',             2014),
        (35, 'Cheap Thrills',               'Sia',                      2016),
        (40, 'One Dance',                   'Drake',                    2016),
        (45, 'Stay With Me',                'Sam Smith',                2014),
        (50, 'Radioactive',                 'Imagine Dragons',          2013),
        (55, 'Royals',                      'Lorde',                    2013),
        (60, 'Counting Stars',              'OneRepublic',              2014),
        (65, 'Let Her Go',                  'Passenger',                2013),
        (70, 'Demons',                      'Imagine Dragons',          2013),
        (75, 'Timber',                      'Pitbull ft. Kesha',        2014),
        (80, 'All of Me',                   'John Legend',              2014),
        (85, 'Maps',                        'Maroon 5',                 2014),
    ]

    for r, (pos, title, artist, year) in enumerate(listened_songs, 2):
        ws.cell(row=r, column=1, value=pos)
        ws.cell(row=r, column=2, value=title)
        ws.cell(row=r, column=3, value=artist)
        ws.cell(row=r, column=4, value=year)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open Chrome to Billboard Hot 100 first
    launch_gui('google-chrome "https://www.billboard.com/charts/hot-100"', delay_sec=3.0)

    # Open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)

    print('GUI_READY: launched Chrome (billboard.com/charts/hot-100) and LibreOffice Calc with DISPLAY=:0')


create_initial()
