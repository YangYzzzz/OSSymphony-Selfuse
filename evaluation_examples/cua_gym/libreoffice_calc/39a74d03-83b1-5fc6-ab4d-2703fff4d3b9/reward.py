"""
Reward Script: Check Billboard Hot 100 top-30 songs released before 2024 not in listening history
Task ID: osworld_multi_apps_misc_014
Domain: libreoffice_calc
Scoring:
  Component 1: 'unheard_songs' sheet exists in the xlsx file (0.30 pts)
  Component 2: Headers of 'unheard_songs' match original sheet columns (0.20 pts)
  Component 3: All rows have Year < 2024 and Position <= 30 (pre-2024 top-30 filter) (0.20 pts)
  Component 4: Rows are sorted by Position ascending (0.15 pts)
  Component 5: No songs from 'my_songs' appear in 'unheard_songs' (0.15 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_misc_014'

# Try both possible filenames (setup may use task_id or original name)
CANDIDATE_FILES = [
    os.path.join(WORKDIR, f'{TASK_ID}.xlsx'),
    os.path.join(WORKDIR, 'my_songs.xlsx'),
]


def find_file():
    """Return the first xlsx file that exists."""
    for path in CANDIDATE_FILES:
        if os.path.exists(path):
            return path
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: 'unheard_songs' sheet exists (0.30 points)
    try:
        if 'unheard_songs' in wb.sheetnames:
            print("PASS: Component 1 — 'unheard_songs' sheet exists (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 — 'unheard_songs' sheet not found. Found sheets: {wb.sheetnames}")
            # Cannot proceed with other checks if the sheet doesn't exist
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws_unheard = wb['unheard_songs']

    # Component 2: Headers of 'unheard_songs' match expected columns (0.20 points)
    # Expected columns from task: Position, Title, Artist, Year
    try:
        header_row = [cell.value for cell in ws_unheard[1]]
        expected_headers = ['Position', 'Title', 'Artist', 'Year']
        # Check all 4 expected headers are present (case-insensitive match)
        actual_lower = [str(h).strip().lower() if h is not None else '' for h in header_row[:4]]
        expected_lower = [h.lower() for h in expected_headers]
        if actual_lower == expected_lower:
            print(f"PASS: Component 2 — Headers match {expected_headers} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 — Headers mismatch. Expected {expected_headers}, found {header_row[:4]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Read all data rows from 'unheard_songs' (skip header)
    data_rows = []
    try:
        for row in ws_unheard.iter_rows(min_row=2, max_row=ws_unheard.max_row, values_only=True):
            if row[0] is not None:  # skip empty rows
                data_rows.append(row)
    except Exception as e:
        print(f"ERROR: Reading data rows — {e}")

    # Component 3: All rows have Year < 2024 and Position <= 30 (0.20 points)
    # This verifies the filtering criteria were applied correctly
    try:
        if len(data_rows) == 0:
            print("FAIL: Component 3 — No data rows found in 'unheard_songs'")
        else:
            violations = []
            for row in data_rows:
                position = row[0]
                year = row[3]
                if position is None or year is None:
                    violations.append(f"Row with None position/year: {row}")
                    continue
                try:
                    pos_int = int(position)
                    year_int = int(year)
                    if pos_int > 30:
                        violations.append(f"Position {pos_int} > 30 (not in top 30)")
                    if year_int >= 2024:
                        violations.append(f"Year {year_int} >= 2024 (not pre-2024)")
                except (ValueError, TypeError) as ve:
                    violations.append(f"Cannot parse position/year from {row}: {ve}")

            if len(violations) == 0:
                print(f"PASS: Component 3 — All {len(data_rows)} rows have Year < 2024 and Position <= 30 (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 3 — Filter criteria violations: {violations[:3]}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Rows are sorted by Position ascending (0.15 points)
    try:
        if len(data_rows) == 0:
            print("FAIL: Component 4 — No data rows to check sorting")
        else:
            positions = []
            for row in data_rows:
                try:
                    positions.append(int(row[0]))
                except (ValueError, TypeError):
                    positions.append(None)

            valid_positions = [p for p in positions if p is not None]
            if valid_positions == sorted(valid_positions):
                print(f"PASS: Component 4 — Rows are sorted by Position ascending: {valid_positions} (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 4 — Rows not sorted by Position. Found: {valid_positions}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: No songs from 'my_songs' appear in 'unheard_songs' (0.15 points)
    # Songs in 'my_songs' are already heard, so they should NOT appear in 'unheard_songs'
    try:
        # Read existing songs from 'my_songs' sheet
        if 'my_songs' not in wb.sheetnames:
            print("FAIL: Component 5 — 'my_songs' sheet not found, cannot verify exclusion")
        else:
            ws_known = wb['my_songs']
            known_titles = set()
            for row in ws_known.iter_rows(min_row=2, max_row=ws_known.max_row, values_only=True):
                if row[1] is not None:
                    known_titles.add(str(row[1]).strip().lower())

            # Check no unheard song title appears in known songs
            duplicates = []
            for row in data_rows:
                if row[1] is not None:
                    title = str(row[1]).strip().lower()
                    if title in known_titles:
                        duplicates.append(row[1])

            if not duplicates:
                print(f"PASS: Component 5 — No songs from 'my_songs' appear in 'unheard_songs' (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 5 — Songs already in 'my_songs' found in 'unheard_songs': {duplicates}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


file_path = find_file()
if file_path is None:
    print(f"File not found. Tried: {CANDIDATE_FILES}")
    print("REWARD: 0.0")
else:
    print(f"Using file: {file_path}")
    verify_task(file_path)
