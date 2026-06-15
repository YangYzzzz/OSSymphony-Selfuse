"""
Initial Setup: Create a spreadsheet with sales rep data, B2 empty with no validation.
Task ID: calc_nrv_056
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_056'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SalesData"

    # --- Headers ---
    headers = ['Sales Rep', 'Region', 'Territory']
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # --- Data rows ---
    # B2 is intentionally left EMPTY (no value, no validation) - that is the task target
    data = [
        ['Elena Rodriguez', '', 'Pacific Northwest'],
        ['Marcus Thompson', 'South', 'Gulf Coast'],
        ['Sarah Chen', 'East', 'New England'],
        ['James Okafor', 'West', 'Mountain States'],
        ['Priya Sharma', 'North', 'Great Lakes'],
        ['David Kim', 'South', 'Southeast'],
        ['Amara Johnson', 'East', 'Mid-Atlantic'],
        ['Lucas Fernandez', 'West', 'Southwest'],
        ['Olivia Brown', 'North', 'Upper Midwest'],
        ['Raj Patel', 'South', 'Gulf States'],
        ['Mia Williams', 'East', 'Tri-State'],
        ['Carlos Gutierrez', 'West', 'Pacific Coast'],
    ]

    data_font = Font(name="Calibri", size=11)
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val if val != '' else None)
            cell.font = data_font
            cell.border = data_border

    # Column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 22

    # --- Second sheet: RegionInfo ---
    ws2 = wb.create_sheet("RegionInfo")
    ws2["A1"] = "Region"
    ws2["B1"] = "Headquarters"
    ws2["C1"] = "Manager"
    region_info = [
        ['North', 'Chicago, IL', 'Tom Bradley'],
        ['South', 'Atlanta, GA', 'Linda Foster'],
        ['East', 'New York, NY', 'Karen White'],
        ['West', 'Denver, CO', 'Mike Santos'],
    ]
    for col in range(1, 4):
        cell = ws2.cell(row=1, column=col)
        cell.font = Font(bold=True)
    for r, row_data in enumerate(region_info, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
