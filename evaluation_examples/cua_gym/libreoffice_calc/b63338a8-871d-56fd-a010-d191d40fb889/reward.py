"""
Reward Script: Class Gradebook with Formulas and Conditional Formatting
Task ID: calc_gen_education_070
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Current Average formulas in J2:J31 (AVERAGE-based)  — 0.25 pts
  Component 2: Letter Grade formulas in K2:K31 (IF-based grading)  — 0.25 pts
  Component 3: Needed-for-A/B formulas in L2:M31                   — 0.25 pts
  Component 4: At Risk flag formulas in N2:N31                      — 0.15 pts
  Component 5: Conditional formatting for AT RISK rows (red fill)   — 0.10 pts
  Total: 1.0
"""

import os
import openpyxl
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_education_070'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that Gradebook sheet exists
    if 'Gradebook' not in wb.sheetnames:
        print("CRITICAL: 'Gradebook' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Gradebook']

    # -----------------------------------------------------------------------
    # Component 1: Current Average formulas in J2:J31 (0.25 points)
    # Task requires each row to have an AVERAGE formula over B:I in column J
    # This FAILS on initial (all None) → should PASS on golden
    # -----------------------------------------------------------------------
    try:
        avg_formula_count = 0
        avg_formula_correct = 0
        for row in range(2, 32):
            val = ws.cell(row, 10).value  # column J
            if val is not None:
                avg_formula_count += 1
                # Accept AVERAGE or IFERROR(AVERAGE(...)) patterns
                val_str = str(val).upper().replace(' ', '')
                if 'AVERAGE' in val_str and f'B{row}' in str(val) and f'I{row}' in str(val):
                    avg_formula_correct += 1

        if avg_formula_correct == 30:
            print(f"PASS: Component 1 — All 30 AVERAGE formulas present in J2:J31 (0.25 pts)")
            total_score += 0.25
        elif avg_formula_correct >= 20:
            partial = round(0.25 * avg_formula_correct / 30, 4)
            print(f"PARTIAL: Component 1 — {avg_formula_correct}/30 AVERAGE formulas in J column ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Expected 30 AVERAGE formulas in J2:J31, found {avg_formula_correct} "
                  f"(total non-None: {avg_formula_count})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Letter Grade formulas in K2:K31 (0.25 points)
    # Task requires IF-based grading in column K referencing column J or grade thresholds
    # This FAILS on initial (all None) → should PASS on golden
    # -----------------------------------------------------------------------
    try:
        grade_formula_count = 0
        for row in range(2, 32):
            val = ws.cell(row, 11).value  # column K
            if val is not None:
                val_str = str(val).upper().replace(' ', '')
                # Must reference column J and use IF or VLOOKUP for grading
                if ('IF' in val_str or 'VLOOKUP' in val_str) and f'J{row}' in str(val):
                    grade_formula_count += 1

        if grade_formula_count == 30:
            print(f"PASS: Component 2 — All 30 letter grade formulas present in K2:K31 (0.25 pts)")
            total_score += 0.25
        elif grade_formula_count >= 20:
            partial = round(0.25 * grade_formula_count / 30, 4)
            print(f"PARTIAL: Component 2 — {grade_formula_count}/30 letter grade formulas in K column ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Expected 30 letter grade formulas in K2:K31, found {grade_formula_count}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Needed-for-A formulas (L2:L31) and Needed-for-B formulas (M2:M31) (0.25 points)
    # Task requires formulas calculating score needed on remaining assignments
    # Pattern: =(8*90 - SUM(B:I)) / (8 - COUNT(B:I)) for A
    #          =(8*80 - SUM(B:I)) / (8 - COUNT(B:I)) for B
    # This FAILS on initial (all None) → should PASS on golden
    # -----------------------------------------------------------------------
    try:
        needed_a_count = 0
        needed_b_count = 0
        for row in range(2, 32):
            val_l = ws.cell(row, 12).value  # column L — Needed for A
            val_m = ws.cell(row, 13).value  # column M — Needed for B

            if val_l is not None:
                l_str = str(val_l).upper().replace(' ', '')
                # Must contain SUM and COUNT referencing B:I
                if ('SUM' in l_str and 'COUNT' in l_str and
                        f'B{row}' in str(val_l) and f'I{row}' in str(val_l)):
                    needed_a_count += 1

            if val_m is not None:
                m_str = str(val_m).upper().replace(' ', '')
                if ('SUM' in m_str and 'COUNT' in m_str and
                        f'B{row}' in str(val_m) and f'I{row}' in str(val_m)):
                    needed_b_count += 1

        # Both L and M columns must be complete
        both_complete = min(needed_a_count, needed_b_count)
        if needed_a_count == 30 and needed_b_count == 30:
            print(f"PASS: Component 3 — All 30 'Needed for A' (L) and 'Needed for B' (M) formulas present (0.25 pts)")
            total_score += 0.25
        elif both_complete >= 20:
            partial = round(0.25 * both_complete / 30, 4)
            print(f"PARTIAL: Component 3 — {needed_a_count}/30 Needed-A, {needed_b_count}/30 Needed-B formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Expected 30 each in L and M columns, found "
                  f"Needed-A={needed_a_count}, Needed-B={needed_b_count}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: At Risk flag formulas in N2:N31 (0.15 points)
    # Task requires IF formula checking if current avg < 60 → "AT RISK", else ""
    # This FAILS on initial (all None) → should PASS on golden
    # -----------------------------------------------------------------------
    try:
        at_risk_formula_count = 0
        for row in range(2, 32):
            val = ws.cell(row, 14).value  # column N
            if val is not None:
                val_str = str(val).upper().replace(' ', '')
                # Must be an IF formula referencing column J (or checking average < 60)
                if 'IF' in val_str and ('60' in val_str or 'ATRISK' in val_str.replace('"', '').replace(' ', '')):
                    at_risk_formula_count += 1

        if at_risk_formula_count == 30:
            print(f"PASS: Component 4 — All 30 'At Risk' flag formulas present in N2:N31 (0.15 pts)")
            total_score += 0.15
        elif at_risk_formula_count >= 20:
            partial = round(0.15 * at_risk_formula_count / 30, 4)
            print(f"PARTIAL: Component 4 — {at_risk_formula_count}/30 'At Risk' formulas in N column ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Expected 30 AT RISK formulas in N2:N31, found {at_risk_formula_count}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -----------------------------------------------------------------------
    # Component 5: Conditional formatting for AT RISK rows (0.10 points)
    # Task requires red highlight on rows where N column = "AT RISK"
    # This FAILS on initial (no CF rules) → should PASS on golden
    # -----------------------------------------------------------------------
    try:
        cf_rules = ws.conditional_formatting
        has_at_risk_cf = False
        has_red_fill = False

        for cf_range in cf_rules:
            for rule in cf_rules[cf_range]:
                # Check for formula-based rule referencing AT RISK
                if hasattr(rule, 'formula') and rule.formula:
                    formula_str = ' '.join(str(f) for f in rule.formula).upper()
                    if 'AT RISK' in formula_str or 'ATRISK' in formula_str.replace(' ', '').replace('"', ''):
                        has_at_risk_cf = True
                        # Check for red fill
                        if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                            fill_color = rule.dxf.fill.fgColor.rgb if rule.dxf.fill.fgColor else None
                            if fill_color:
                                # Check for red-ish color (FF in the R channel with minimal G/B)
                                # Common red: FFFF0000, FFFF9999, FFE26B0A, etc.
                                try:
                                    r = int(fill_color[2:4], 16)
                                    g = int(fill_color[4:6], 16)
                                    b = int(fill_color[6:8], 16)
                                    # Red: R is dominant
                                    if r > 150 and r > g and r > b:
                                        has_red_fill = True
                                except (ValueError, IndexError):
                                    pass

        if has_at_risk_cf and has_red_fill:
            print(f"PASS: Component 5 — Conditional formatting with red fill for AT RISK rows found (0.10 pts)")
            total_score += 0.10
        elif has_at_risk_cf:
            print(f"PARTIAL: Component 5 — AT RISK conditional formatting found but red fill not confirmed (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — No conditional formatting found for AT RISK rows")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
