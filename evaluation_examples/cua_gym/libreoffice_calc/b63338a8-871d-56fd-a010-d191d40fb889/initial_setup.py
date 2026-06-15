"""
Initial Setup: Class gradebook with 30 students and 8 assignments
Task ID: calc_gen_education_070
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_education_070'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ #
    # Sheet 1: Gradebook
    # ------------------------------------------------------------------ #
    ws = wb.active
    ws.title = 'Gradebook'

    # Headers
    headers = [
        'Student', 'A1', 'A2', 'A3', 'A4', 'A5', 'A6', 'A7', 'A8',
        'Current Avg', 'Letter Grade', 'Needed for A', 'Needed for B', 'At Risk'
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.alignment = Alignment(horizontal='center')

    # 30 students — realistic names, partial assignment scores
    # Some assignments are blank (not yet graded)
    students_data = [
        # Name,          A1,    A2,    A3,    A4,    A5,    A6,    A7,    A8
        ['Emma Hartley',    88,    92,    85,    90,    78,    None,  None,  None],
        ['Liam Nguyen',     75,    80,    70,    None,  None,  None,  None,  None],
        ['Olivia Chen',     95,    98,    92,    96,    94,    91,    97,    93],
        ['Noah Patel',      62,    55,    68,    71,    None,  None,  None,  None],
        ['Ava Robinson',    85,    88,    90,    87,    82,    86,    None,  None],
        ['Ethan Williams',  72,    65,    78,    80,    None,  None,  None,  None],
        ['Sophia Martinez', 91,    94,    89,    95,    92,    88,    96,    90],
        ['Mason Lee',       58,    63,    55,    60,    None,  None,  None,  None],
        ['Isabella Davis',  79,    82,    85,    88,    76,    None,  None,  None],
        ['Logan Thomas',    84,    87,    80,    83,    89,    85,    None,  None],
        ['Mia Garcia',      96,    99,    97,    100,   95,    98,    94,    97],
        ['Lucas Anderson',  67,    70,    65,    None,  None,  None,  None,  None],
        ['Charlotte Wilson', 88,   90,    84,    92,    87,    None,  None,  None],
        ['Aiden Jackson',   73,    68,    75,    77,    None,  None,  None,  None],
        ['Amelia White',    90,    93,    88,    95,    91,    89,    92,    None],
        ['Elijah Harris',   55,    58,    50,    62,    None,  None,  None,  None],
        ['Harper Thompson', 82,    85,    79,    88,    83,    None,  None,  None],
        ['James Brown',     76,    72,    80,    None,  None,  None,  None,  None],
        ['Evelyn Taylor',   93,    91,    95,    97,    90,    94,    88,    96],
        ['Alexander Moore', 61,    65,    58,    70,    None,  None,  None,  None],
        ['Abigail Martin',  87,    89,    83,    91,    86,    88,    None,  None],
        ['Michael Clark',   69,    74,    72,    None,  None,  None,  None,  None],
        ['Emily Rodriguez', 94,    97,    91,    98,    93,    95,    99,    92],
        ['Benjamin Lewis',  78,    75,    82,    79,    None,  None,  None,  None],
        ['Avery Walker',    86,    90,    88,    84,    91,    87,    None,  None],
        ['Daniel Hall',     64,    60,    67,    None,  None,  None,  None,  None],
        ['Sofia Young',     92,    95,    90,    96,    93,    91,    94,    None],
        ['Matthew Allen',   70,    73,    68,    75,    None,  None,  None,  None],
        ['Scarlett King',   83,    86,    81,    89,    84,    80,    None,  None],
        ['Jackson Scott',   57,    52,    61,    None,  None,  None,  None,  None],
    ]

    for r, row_data in enumerate(students_data, 2):
        for c, val in enumerate(row_data, 1):
            if val is not None:
                ws.cell(row=r, column=c, value=val)
        # Columns J (10), K (11), L (12), M (13), N (14) intentionally left EMPTY
        # These are filled in by the agent

    # Column widths
    ws.column_dimensions['A'].width = 22
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col_letter].width = 8
    ws.column_dimensions['J'].width = 14
    ws.column_dimensions['K'].width = 14
    ws.column_dimensions['L'].width = 14
    ws.column_dimensions['M'].width = 14
    ws.column_dimensions['N'].width = 12

    # Freeze header row
    ws.freeze_panes = 'A2'

    # ------------------------------------------------------------------ #
    # Sheet 2: GradeScale
    # ------------------------------------------------------------------ #
    ws2 = wb.create_sheet('GradeScale')

    # Headers
    ws2.cell(row=1, column=1, value='Letter').font = Font(bold=True)
    ws2.cell(row=1, column=2, value='Min Score').font = Font(bold=True)

    # Grade cutoffs: A2:B6
    grade_scale = [
        ['A', 90],
        ['B', 80],
        ['C', 70],
        ['D', 60],
        ['F', 0],
    ]
    for r, (letter, cutoff) in enumerate(grade_scale, 2):
        ws2.cell(row=r, column=1, value=letter)
        ws2.cell(row=r, column=2, value=cutoff)

    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Gradebook sheet: 30 students, 8 assignment columns (partially filled)')
    print(f'  GradeScale sheet: grade cutoffs A2:B6')
    print(f'  Columns J-N (Current Avg, Letter Grade, Needed for A/B, At Risk): EMPTY')


create_initial()
