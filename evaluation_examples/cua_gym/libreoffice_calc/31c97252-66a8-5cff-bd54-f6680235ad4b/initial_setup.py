"""
Initial Setup: Pivot table with blank category rows
Task ID: calc_pivot_051
Domain: libreoffice_calc

Creates a spreadsheet with:
- RawData sheet: 180 rows of ID, Category, Amount (some blanks in Category)
- PivotSheet: A summary pivot-like table showing Category + SUM(Amount), including (blank) row
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_051'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()

    # --- Sheet 1: RawData ---
    ws = wb.active
    ws.title = 'RawData'

    # Headers
    headers = ['ID', 'Category', 'Amount']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align

    # Generate 180 rows of data
    categories = ['Electronics', 'Clothing', 'Food']
    # Decide which rows will have blank category (~12 blanks)
    blank_rows = set(random.sample(range(2, 182), 12))

    # We need to hit specific totals:
    # Electronics=52000, Clothing=38000, Food=28000, blank total = some amount
    # Total with blanks will be 118000 + blank_total

    # Distribute amounts to hit targets
    elec_rows = []
    cloth_rows = []
    food_rows = []
    blank_row_list = []

    for r in range(2, 182):
        if r in blank_rows:
            blank_row_list.append(r)
        else:
            cat_choice = random.choice(categories)
            if cat_choice == 'Electronics':
                elec_rows.append(r)
            elif cat_choice == 'Clothing':
                cloth_rows.append(r)
            else:
                food_rows.append(r)

    def distribute_amounts(rows, target_total):
        """Distribute target_total across rows with realistic amounts."""
        n = len(rows)
        if n == 0:
            return {}
        base = target_total / n
        amounts = {}
        remaining = target_total
        for i, r in enumerate(rows):
            if i == n - 1:
                amounts[r] = round(remaining, 2)
            else:
                # Add some randomness
                amt = round(base * random.uniform(0.4, 1.6), 2)
                amt = min(amt, remaining - (n - i - 1) * 50)  # ensure remaining rows can get at least 50
                amt = max(amt, 50)
                amounts[r] = amt
                remaining -= amt
        return amounts

    elec_amounts = distribute_amounts(elec_rows, 52000)
    cloth_amounts = distribute_amounts(cloth_rows, 38000)
    food_amounts = distribute_amounts(food_rows, 28000)

    # Blank rows get some amounts too
    blank_total = 0
    blank_amounts = {}
    for r in blank_row_list:
        amt = round(random.uniform(200, 1500), 2)
        blank_amounts[r] = amt
        blank_total += amt

    # Write data
    for r in range(2, 182):
        row_id = r - 1
        ws.cell(row=r, column=1, value=row_id)

        if r in blank_rows:
            # Leave Category blank
            ws.cell(row=r, column=2, value=None)
            ws.cell(row=r, column=3, value=blank_amounts[r])
        elif r in elec_amounts:
            ws.cell(row=r, column=2, value='Electronics')
            ws.cell(row=r, column=3, value=elec_amounts[r])
        elif r in cloth_amounts:
            ws.cell(row=r, column=2, value='Clothing')
            ws.cell(row=r, column=3, value=cloth_amounts[r])
        elif r in food_amounts:
            ws.cell(row=r, column=2, value='Food')
            ws.cell(row=r, column=3, value=food_amounts[r])

    # Format Amount column as currency
    for r in range(2, 182):
        ws.cell(row=r, column=3).number_format = '#,##0.00'

    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 14

    # Freeze header
    ws.freeze_panes = 'A2'

    # --- Sheet 2: PivotSheet (simulated pivot table) ---
    ps = wb.create_sheet('PivotSheet')

    # Title
    ps.merge_cells('A1:C1')
    title_cell = ps['A1']
    title_cell.value = 'Pivot Table - Category Summary'
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center")

    # Pivot headers
    pivot_headers = ['Category', 'Sum of Amount']
    pivot_header_fill = PatternFill(start_color="FF305496", end_color="FF305496", fill_type="solid")
    for col, h in enumerate(pivot_headers, 1):
        cell = ps.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = pivot_header_fill
        cell.alignment = Alignment(horizontal="center")

    # Pivot data rows (including blank row)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    pivot_data = [
        ('Clothing', 38000),
        ('Electronics', 52000),
        ('Food', 28000),
        ('(blank)', round(blank_total, 2)),
    ]

    for i, (cat, amt) in enumerate(pivot_data):
        row = 4 + i
        cat_cell = ps.cell(row=row, column=1, value=cat)
        amt_cell = ps.cell(row=row, column=2, value=amt)
        amt_cell.number_format = '#,##0.00'
        cat_cell.border = thin_border
        amt_cell.border = thin_border
        if cat == '(blank)':
            cat_cell.font = Font(italic=True, color="888888")

    # Grand Total row
    grand_total = 52000 + 38000 + 28000 + round(blank_total, 2)
    gt_row = 4 + len(pivot_data)
    gt_cat = ps.cell(row=gt_row, column=1, value='Grand Total')
    gt_cat.font = Font(bold=True)
    gt_cat.border = thin_border
    gt_amt = ps.cell(row=gt_row, column=2, value=grand_total)
    gt_amt.font = Font(bold=True)
    gt_amt.number_format = '#,##0.00'
    gt_amt.border = thin_border

    # Column widths
    ps.column_dimensions['A'].width = 20
    ps.column_dimensions['B'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  RawData: 180 rows, {len(blank_row_list)} blank categories')
    print(f'  PivotSheet: 4 categories (including blank), Grand Total = {grand_total}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
