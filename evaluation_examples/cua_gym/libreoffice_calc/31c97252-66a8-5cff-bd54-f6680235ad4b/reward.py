"""
Reward Script: Remove blank rows from pivot table display
Task ID: calc_pivot_051
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35) - No "(blank)" category row exists in PivotSheet
  Component 2 (0.35) - Grand Total reflects only non-blank categories (118000)
  Component 3 (0.30) - Exactly 3 category rows (Clothing, Electronics, Food) present
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_051'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: PivotSheet must exist
    if 'PivotSheet' not in wb.sheetnames:
        print("CRITICAL: 'PivotSheet' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['PivotSheet']

    # Scan all rows in column A to find category entries and detect "(blank)"
    # We look for all non-empty values in column A below the header row
    all_data_rows = []  # all data rows (including blank entries)
    blank_row_found = False
    grand_total_value = None

    for row_idx in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row_idx, column=1).value
        cell_b = ws.cell(row=row_idx, column=2).value
        if cell_a is None:
            continue
        cell_a_str = str(cell_a).strip()
        # Detect blank row - could be "(blank)", "blank", "(Blank)", etc.
        if cell_a_str.lower().replace('(', '').replace(')', '') == 'blank':
            blank_row_found = True
            all_data_rows.append(cell_a_str)
        elif cell_a_str == 'Grand Total':
            grand_total_value = cell_b
        elif cell_a_str in ('Category', 'Pivot Table - Category Summary'):
            # Header rows, skip
            continue
        else:
            all_data_rows.append(cell_a_str)

    print(f"DEBUG: all_data_rows = {all_data_rows}")
    print(f"DEBUG: blank_row_found = {blank_row_found}")
    print(f"DEBUG: grand_total_value = {grand_total_value}")

    # Component 1: No "(blank)" category row exists in PivotSheet (0.35 points)
    # This FAILS on initial (has blank row) -> PASSES on golden (no blank row)
    try:
        if not blank_row_found:
            print(f"PASS: Component 1 - No '(blank)' row found in pivot table (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 1 - '(blank)' row still present in pivot table")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Grand Total reflects only non-blank categories (~118000) (0.35 points)
    # Initial has 127786.45, golden should have 118000
    # This FAILS on initial -> PASSES on golden
    try:
        if grand_total_value is not None:
            gt = float(grand_total_value)
            # The expected grand total is 118000 (sum of Electronics=52000 + Clothing=38000 + Food=28000)
            # Allow some tolerance for rounding
            if abs(gt - 118000) < 500:
                print(f"PASS: Component 2 - Grand Total is {gt}, close to expected 118000 (0.35 pts)")
                total_score += 0.35
            else:
                print(f"FAIL: Component 2 - Grand Total is {gt}, expected ~118000")
        else:
            print(f"FAIL: Component 2 - No Grand Total row found")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Exactly 3 data rows in pivot (Clothing, Electronics, Food) - no extra rows (0.30 points)
    # Initial has 4 data rows (including blank), golden should have exactly 3
    # This FAILS on initial -> PASSES on golden
    try:
        expected_categories = {'Clothing', 'Electronics', 'Food'}
        actual_set = set(all_data_rows)
        if actual_set == expected_categories and len(all_data_rows) == 3:
            print(f"PASS: Component 3 - Exactly 3 data rows found: {all_data_rows} (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 3 - Expected exactly 3 rows {{Clothing, Electronics, Food}}, found {len(all_data_rows)} rows: {all_data_rows}")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
