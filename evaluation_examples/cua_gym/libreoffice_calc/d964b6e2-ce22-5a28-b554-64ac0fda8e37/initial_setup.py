"""
Initial Setup: Staff data spreadsheet for pivot table task
Task ID: calc_pivot_006
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_006'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

random.seed(42)

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'StaffData'

    # --- Headers ---
    headers = ['ID', 'Name', 'Department', 'Position', 'Salary', 'YearsExp']
    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    # --- Name pools ---
    first_names = [
        'Sarah', 'Marcus', 'Elena', 'James', 'Priya', 'David', 'Mei',
        'Carlos', 'Aisha', 'Robert', 'Yuki', 'Thomas', 'Fatima', 'Kevin',
        'Olga', 'Daniel', 'Nadia', 'Brian', 'Lena', 'Victor', 'Amara',
        'Patrick', 'Sofia', 'Andre', 'Hannah', 'Omar', 'Julia', 'Nathan',
        'Rina', 'George', 'Chloe', 'Ivan', 'Mira', 'Samuel', 'Tara',
        'Felix', 'Diana', 'Leo', 'Vera', 'Hugo', 'Zara', 'Peter',
        'Lily', 'Max', 'Rosa',
    ]
    last_names = [
        'Chen', 'Johnson', 'Petrov', 'Williams', 'Sharma', 'Kim', 'Garcia',
        'Mueller', 'Okafor', 'Smith', 'Tanaka', 'Brown', 'Hassan', 'Lee',
        'Ivanova', 'Martinez', 'Becker', 'Clark', 'Novak', 'Wright',
        'Diaz', 'Fischer', 'Ali', 'Taylor', 'Nakamura', 'Anderson',
        'Johansson', 'Patel', 'Wilson', 'Costa', 'Thompson', 'Berg',
        'Morales', 'Reed', 'Kowalski', 'Park', 'Evans', 'Larsson',
        'Sato', 'Hughes', 'Ferreira', 'Grant', 'Rossi', 'Lambert', 'Cruz',
    ]

    departments = ['IT', 'HR', 'Sales', 'Operations', 'Legal']

    positions_by_dept = {
        'IT': ['Software Engineer', 'Data Analyst', 'DevOps Engineer', 'QA Tester', 'IT Manager', 'System Admin'],
        'HR': ['HR Specialist', 'Recruiter', 'HR Manager', 'Training Coordinator', 'Benefits Analyst'],
        'Sales': ['Sales Rep', 'Account Manager', 'Sales Director', 'Business Dev', 'Sales Analyst'],
        'Operations': ['Operations Analyst', 'Logistics Coordinator', 'Project Manager', 'Process Engineer', 'Supply Chain Mgr'],
        'Legal': ['Legal Counsel', 'Paralegal', 'Compliance Officer', 'Contract Specialist', 'Legal Director'],
    }

    # Ground truth max salaries: IT=115000, HR=92000, Sales=105000, Operations=88000, Legal=120000
    # We need to ensure exactly these maxima per department.
    # Strategy: assign 18 employees per department, set one employee to the max salary,
    # and the rest below it.

    max_salaries = {
        'IT': 115000,
        'HR': 92000,
        'Sales': 105000,
        'Operations': 88000,
        'Legal': 120000,
    }

    salary_ranges = {
        'IT': (55000, 114000),
        'HR': (40000, 91000),
        'Sales': (45000, 104000),
        'Operations': (38000, 87000),
        'Legal': (50000, 119000),
    }

    # Build 90 rows: 18 per department
    rows = []
    emp_id = 1
    used_names = set()

    for dept in departments:
        lo, hi = salary_ranges[dept]
        max_sal = max_salaries[dept]
        positions = positions_by_dept[dept]

        for i in range(18):
            # Pick a unique name
            while True:
                fn = random.choice(first_names)
                ln = random.choice(last_names)
                full = f'{fn} {ln}'
                if full not in used_names:
                    used_names.add(full)
                    break

            pos = positions[i % len(positions)]
            yrs = random.randint(1, 25)

            if i == 0:
                salary = max_sal  # ensure max is hit
            else:
                salary = random.randint(lo // 1000, hi // 1000) * 1000

            rows.append([emp_id, full, dept, pos, salary, yrs])
            emp_id += 1

    # Shuffle so departments are interleaved (more realistic)
    random.shuffle(rows)

    # Re-assign sequential IDs after shuffle
    for idx, row in enumerate(rows):
        row[0] = idx + 1

    # Write data rows
    for r, row_data in enumerate(rows, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 24
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12

    # Number format for salary
    for r in range(2, 92):
        ws.cell(row=r, column=5).number_format = '#,##0'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
