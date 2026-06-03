"""
Reward Script: Generate a pivot table showing maximum salary per department
Task ID: calc_pivot_006
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): Pivot sheet exists with correct headers
  Component 2 (0.5): All 5 department max salary values are correct
  Component 3 (0.2): Exactly 5 departments, no extra/missing rows
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_006'

# Expected max salaries per department (from task context ground truth)
EXPECTED = {
    'IT': 115000,
    'HR': 92000,
    'Sales': 105000,
    'Operations': 88000,
    'Legal': 120000,
}


def persist_app_state():
    """Save any unsaved LibreOffice edits before verification."""
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(1.0)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Pivot sheet exists with correct headers (0.3 points)
    # This FAILS on initial (no Pivot sheet) and PASSES on golden
    try:
        pivot_sheet = None
        for name in wb.sheetnames:
            if name.lower().strip() in ('pivot', 'pivot table', 'pivottable'):
                pivot_sheet = wb[name]
                break

        if pivot_sheet is None:
            print(f"FAIL: Component 1 — No pivot sheet found. Sheets: {wb.sheetnames}")
        else:
            header_a = str(pivot_sheet.cell(row=1, column=1).value or '').strip().lower()
            header_b = str(pivot_sheet.cell(row=1, column=2).value or '').strip().lower()

            has_dept_header = 'department' in header_a
            has_salary_header = 'max' in header_b and 'salary' in header_b

            if has_dept_header and has_salary_header:
                print(f"PASS: Component 1 — Pivot sheet found with correct headers (0.3 pts)")
                total_score += 0.3
            elif has_dept_header or has_salary_header:
                print(f"PARTIAL: Component 1 — Pivot sheet found, partial headers: A1='{header_a}', B1='{header_b}' (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 1 — Pivot sheet headers incorrect: A1='{header_a}', B1='{header_b}'")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # If no pivot sheet found, remaining checks cannot proceed
    if pivot_sheet is None:
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    # Read pivot data from the sheet (rows 2 onward, columns A and B)
    pivot_data = {}
    try:
        for row_idx in range(2, pivot_sheet.max_row + 1):
            dept_val = pivot_sheet.cell(row=row_idx, column=1).value
            salary_val = pivot_sheet.cell(row=row_idx, column=2).value
            if dept_val is not None and salary_val is not None:
                dept_key = str(dept_val).strip()
                try:
                    salary_num = float(salary_val)
                except (ValueError, TypeError):
                    salary_num = None
                if dept_key and salary_num is not None:
                    pivot_data[dept_key] = salary_num
        print(f"INFO: Pivot data read: {pivot_data}")
    except Exception as e:
        print(f"ERROR: Reading pivot data — {e}")

    # Component 2: All 5 department max salary values are correct (0.5 points)
    # Each correct value = 0.1 points (5 * 0.1 = 0.5)
    try:
        correct_count = 0
        for dept, expected_salary in EXPECTED.items():
            if dept in pivot_data:
                actual = pivot_data[dept]
                if abs(actual - expected_salary) < 1.0:
                    print(f"PASS: Component 2 — {dept} max salary = {actual} (expected {expected_salary})")
                    correct_count += 1
                else:
                    print(f"FAIL: Component 2 — {dept} max salary = {actual}, expected {expected_salary}")
            else:
                print(f"FAIL: Component 2 — Department '{dept}' not found in pivot data")

        if correct_count > 0:
            pts = correct_count * 0.1
            print(f"PASS: Component 2 — {correct_count}/5 departments correct ({pts} pts)")
            total_score += pts
        else:
            print(f"FAIL: Component 2 — No departments matched expected values")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Exactly 5 departments present (0.2 points)
    # Verifies completeness — no extra or missing departments
    try:
        actual_depts = set(pivot_data.keys())
        expected_depts = set(EXPECTED.keys())

        if actual_depts == expected_depts:
            print(f"PASS: Component 3 — Exactly 5 expected departments present (0.2 pts)")
            total_score += 0.2
        else:
            missing = expected_depts - actual_depts
            extra = actual_depts - expected_depts
            print(f"FAIL: Component 3 — Department mismatch. Missing: {missing}, Extra: {extra}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state()

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
