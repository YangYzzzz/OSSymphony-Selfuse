"""
Initial Setup: Data table with broken SUM formula after row insertion
Task ID: calc_tbl_057
Domain: libreoffice_calc

Creates a Revenue Data spreadsheet with 54 data rows (rows 2-55) and a totaling
formula at B56 that still references the OLD range =SUM(B2:B50), missing the 5
rows inserted between the original rows 25-29.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_057'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revenue Data"

    # --- Header row ---
    headers = ["Description", "Amount ($)", "Category", "Date", "Status"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Data rows (54 rows: rows 2 through 55) ---
    # Simulating a monthly revenue table with various line items
    categories = ["Product Sales", "Service Revenue", "Licensing", "Consulting", "Maintenance"]
    statuses = ["Confirmed", "Pending", "Invoiced", "Received", "Projected"]

    descriptions = [
        "Q1 North Region - Enterprise Licenses",
        "Q1 North Region - Support Contracts",
        "Q1 South Region - Enterprise Licenses",
        "Q1 South Region - Hardware Bundles",
        "Q1 West Region - Consulting Services",
        "Q1 West Region - Training Programs",
        "Q1 East Region - SaaS Subscriptions",
        "Q1 East Region - Custom Development",
        "Q1 Central Region - Annual Renewals",
        "Q1 Central Region - Professional Services",
        "Q1 Online Channel - Digital Products",
        "Q1 Online Channel - Marketplace Fees",
        "Q2 North Region - Enterprise Licenses",
        "Q2 North Region - Support Contracts",
        "Q2 South Region - Enterprise Licenses",
        "Q2 South Region - Hardware Bundles",
        "Q2 West Region - Consulting Services",
        "Q2 West Region - Training Programs",
        "Q2 East Region - SaaS Subscriptions",
        "Q2 East Region - Custom Development",
        "Q2 Central Region - Annual Renewals",
        "Q2 Central Region - Professional Services",
        "Q2 Online Channel - Digital Products",
        "Q2 Online Channel - Marketplace Fees",
        # Rows 26-30: the 5 "newly inserted" rows (originally between old rows 25-29)
        "Q2 Partner Channel - Reseller Revenue",
        "Q2 Partner Channel - Affiliate Commissions",
        "Q2 Partner Channel - OEM Licensing",
        "Q2 Partner Channel - Co-Marketing Revenue",
        "Q2 Partner Channel - Integration Fees",
        # Continue with Q3/Q4 data
        "Q3 North Region - Enterprise Licenses",
        "Q3 North Region - Support Contracts",
        "Q3 South Region - Enterprise Licenses",
        "Q3 South Region - Hardware Bundles",
        "Q3 West Region - Consulting Services",
        "Q3 West Region - Training Programs",
        "Q3 East Region - SaaS Subscriptions",
        "Q3 East Region - Custom Development",
        "Q3 Central Region - Annual Renewals",
        "Q3 Central Region - Professional Services",
        "Q3 Online Channel - Digital Products",
        "Q3 Online Channel - Marketplace Fees",
        "Q4 North Region - Enterprise Licenses",
        "Q4 North Region - Support Contracts",
        "Q4 South Region - Enterprise Licenses",
        "Q4 South Region - Hardware Bundles",
        "Q4 West Region - Consulting Services",
        "Q4 West Region - Training Programs",
        "Q4 East Region - SaaS Subscriptions",
        "Q4 East Region - Custom Development",
        "Q4 Central Region - Annual Renewals",
        "Q4 Central Region - Professional Services",
        "Q4 Online Channel - Digital Products",
        "Q4 Online Channel - Marketplace Fees",
        "Q4 Partner Channel - Year-End Bonuses",
    ]

    amounts = [
        45230.00, 12800.00, 38750.00, 22100.00, 31500.00,
        8900.00, 56200.00, 41300.00, 27650.00, 19400.00,
        33100.00, 7250.00, 48500.00, 13200.00, 41900.00,
        24300.00, 35800.00, 9600.00, 59100.00, 44200.00,
        29100.00, 21500.00, 36400.00, 8100.00,
        # 5 inserted rows
        15600.00, 4200.00, 18900.00, 7350.00, 11200.00,
        # Q3/Q4
        51200.00, 14100.00, 43800.00, 25600.00, 37200.00,
        10400.00, 62300.00, 47500.00, 31200.00, 22800.00,
        38500.00, 8700.00, 54600.00, 15300.00, 46100.00,
        27400.00, 39500.00, 11200.00, 64800.00, 49700.00,
        33600.00, 24100.00, 40200.00, 9500.00,
        12750.00,
    ]

    dates = [
        "2025-01-15", "2025-01-22", "2025-02-03", "2025-02-10", "2025-02-18",
        "2025-02-25", "2025-03-05", "2025-03-12", "2025-03-19", "2025-03-26",
        "2025-03-28", "2025-03-31", "2025-04-08", "2025-04-15", "2025-05-02",
        "2025-05-10", "2025-05-18", "2025-05-25", "2025-06-03", "2025-06-12",
        "2025-06-19", "2025-06-25", "2025-06-28", "2025-06-30",
        # 5 inserted rows
        "2025-06-15", "2025-06-17", "2025-06-20", "2025-06-22", "2025-06-24",
        # Q3/Q4
        "2025-07-07", "2025-07-15", "2025-08-01", "2025-08-10", "2025-08-18",
        "2025-08-25", "2025-09-03", "2025-09-12", "2025-09-19", "2025-09-26",
        "2025-09-28", "2025-09-30", "2025-10-08", "2025-10-15", "2025-11-02",
        "2025-11-10", "2025-11-18", "2025-11-25", "2025-12-03", "2025-12-12",
        "2025-12-19", "2025-12-25", "2025-12-28", "2025-12-31",
        "2025-12-30",
    ]

    currency_fmt = '$#,##0.00'
    data_font = Font(name="Calibri", size=11)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    # Alternating row fill
    light_fill = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")

    for i in range(54):
        row = i + 2
        cat_idx = i % len(categories)
        stat_idx = i % len(statuses)

        ws.cell(row=row, column=1, value=descriptions[i]).font = data_font
        amt_cell = ws.cell(row=row, column=2, value=amounts[i])
        amt_cell.font = data_font
        amt_cell.number_format = currency_fmt
        ws.cell(row=row, column=3, value=categories[cat_idx]).font = data_font
        ws.cell(row=row, column=4, value=dates[i]).font = data_font
        ws.cell(row=row, column=5, value=statuses[stat_idx]).font = data_font

        # Apply alternating row shading and borders
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if i % 2 == 1:
                cell.fill = light_fill

    # --- Totals row at row 56 ---
    # Label
    total_font = Font(name="Calibri", size=11, bold=True)
    total_fill = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")

    label_cell = ws.cell(row=56, column=1, value="TOTAL")
    label_cell.font = total_font
    label_cell.fill = total_fill

    # BROKEN formula: still references old range B2:B50, missing rows 51-55
    formula_cell = ws.cell(row=56, column=2, value="=SUM(B2:B50)")
    formula_cell.font = total_font
    formula_cell.fill = total_fill
    formula_cell.number_format = currency_fmt

    for col in range(3, 6):
        cell = ws.cell(row=56, column=col)
        cell.fill = total_fill

    # --- Column widths ---
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
