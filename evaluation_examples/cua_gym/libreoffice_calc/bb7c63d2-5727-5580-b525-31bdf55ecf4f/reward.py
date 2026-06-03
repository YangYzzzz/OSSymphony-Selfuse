"""
Reward Script: Seasonality-adjusted sales forecast with quarterly seasonal indices
Task ID: calc_sales_forecast_seasonal_041
Domain: libreoffice_calc
Scoring:
  - Component 1: Quarterly average formulas in E2:E5 (0.30 pts)
  - Component 2: Overall average formula in F2:F5 (0.20 pts)
  - Component 3: Seasonal index formulas in G2:G5 (0.25 pts)
  - Component 4: Raw trend forecast values in B9:B12 (0.10 pts)
  - Component 5: Seasonality-adjusted forecast formulas in C9:C12 (0.15 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_forecast_seasonal_041'


def normalize_formula(f):
    """Normalize formula string for comparison: uppercase, remove spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook (formula mode)
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists
    if 'SeasonalForecast' not in wb.sheetnames:
        print("CRITICAL: Sheet 'SeasonalForecast' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['SeasonalForecast']

    # Component 1: Quarterly average formulas in E2:E5 (0.30 points)
    # Task requires =AVERAGE(B2:D2), =AVERAGE(B3:D3), =AVERAGE(B4:D4), =AVERAGE(B5:D5)
    try:
        expected_avg_formulas = [
            '=AVERAGE(B2:D2)',
            '=AVERAGE(B3:D3)',
            '=AVERAGE(B4:D4)',
            '=AVERAGE(B5:D5)',
        ]
        coords = ['E2', 'E3', 'E4', 'E5']
        passing = 0
        for coord, expected in zip(coords, expected_avg_formulas):
            val = ws[coord].value
            if isinstance(val, str) and normalize_formula(val) == normalize_formula(expected):
                passing += 1
            else:
                print(f"FAIL Component 1: {coord} expected formula {expected!r}, found {val!r}")
        if passing == 4:
            print(f"PASS: Component 1 — All 4 quarterly average formulas in E2:E5 correct (0.30 pts)")
            total_score += 0.30
        elif passing >= 2:
            partial = round(0.30 * passing / 4, 4)
            print(f"PARTIAL: Component 1 — {passing}/4 quarterly average formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {passing}/4 quarterly average formulas present")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Overall average formula in F2:F5 (0.20 points)
    # Task requires =AVERAGE($E$2:$E$5) in all four cells
    try:
        expected_overall = '=AVERAGE($E$2:$E$5)'
        coords_f = ['F2', 'F3', 'F4', 'F5']
        passing_f = 0
        for coord in coords_f:
            val = ws[coord].value
            if isinstance(val, str) and normalize_formula(val) == normalize_formula(expected_overall):
                passing_f += 1
            else:
                print(f"FAIL Component 2: {coord} expected {expected_overall!r}, found {val!r}")
        if passing_f == 4:
            print(f"PASS: Component 2 — All 4 overall average formulas in F2:F5 correct (0.20 pts)")
            total_score += 0.20
        elif passing_f >= 2:
            partial = round(0.20 * passing_f / 4, 4)
            print(f"PARTIAL: Component 2 — {passing_f}/4 overall average formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {passing_f}/4 overall average formulas in F2:F5 present")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Seasonal index formulas in G2:G5 (0.25 points)
    # Task requires =E2/F2, =E3/F3, =E4/F4, =E5/F5
    try:
        expected_idx_formulas = [
            '=E2/F2',
            '=E3/F3',
            '=E4/F4',
            '=E5/F5',
        ]
        coords_g = ['G2', 'G3', 'G4', 'G5']
        passing_g = 0
        for coord, expected in zip(coords_g, expected_idx_formulas):
            val = ws[coord].value
            if isinstance(val, str) and normalize_formula(val) == normalize_formula(expected):
                passing_g += 1
            else:
                print(f"FAIL Component 3: {coord} expected formula {expected!r}, found {val!r}")
        if passing_g == 4:
            print(f"PASS: Component 3 — All 4 seasonal index formulas in G2:G5 correct (0.25 pts)")
            total_score += 0.25
        elif passing_g >= 2:
            partial = round(0.25 * passing_g / 4, 4)
            print(f"PARTIAL: Component 3 — {passing_g}/4 seasonal index formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {passing_g}/4 seasonal index formulas in G2:G5 present")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Raw trend forecast values in B9:B12 (0.10 points)
    # Task requires non-empty numeric values representing trend forecasts
    # Golden values: B9=2700000, B10=3200000, B11=3100000, B12=4900000
    try:
        coords_b = ['B9', 'B10', 'B11', 'B12']
        passing_b = 0
        for coord in coords_b:
            val = ws[coord].value
            # Check that a numeric trend forecast value exists (non-empty, numeric)
            if val is not None and isinstance(val, (int, float)) and val > 0:
                passing_b += 1
            else:
                print(f"FAIL Component 4: {coord} expected numeric trend forecast value, found {val!r}")
        if passing_b == 4:
            print(f"PASS: Component 4 — All 4 raw trend forecast values in B9:B12 present (0.10 pts)")
            total_score += 0.10
        elif passing_b >= 2:
            partial = round(0.10 * passing_b / 4, 4)
            print(f"PARTIAL: Component 4 — {passing_b}/4 raw trend forecast values present ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {passing_b}/4 raw trend forecast values in B9:B12 present")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Seasonality-adjusted forecast formulas in C9:C12 (0.15 points)
    # Task requires =B9*G2, =B10*G3, =B11*G4, =B12*G5 (raw trend * seasonal index)
    try:
        expected_adj_formulas = [
            '=B9*G2',
            '=B10*G3',
            '=B11*G4',
            '=B12*G5',
        ]
        coords_c = ['C9', 'C10', 'C11', 'C12']
        passing_c = 0
        for coord, expected in zip(coords_c, expected_adj_formulas):
            val = ws[coord].value
            if isinstance(val, str) and normalize_formula(val) == normalize_formula(expected):
                passing_c += 1
            else:
                print(f"FAIL Component 5: {coord} expected formula {expected!r}, found {val!r}")
        if passing_c == 4:
            print(f"PASS: Component 5 — All 4 seasonality-adjusted forecast formulas in C9:C12 correct (0.15 pts)")
            total_score += 0.15
        elif passing_c >= 2:
            partial = round(0.15 * passing_c / 4, 4)
            print(f"PARTIAL: Component 5 — {passing_c}/4 adjusted forecast formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — Only {passing_c}/4 adjusted forecast formulas in C9:C12 present")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
