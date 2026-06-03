"""
Initial Setup: Seasonal sales forecast with 3 years of historical data
Task ID: calc_sales_forecast_seasonal_041
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_forecast_seasonal_041'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- SeasonalForecast sheet ---
    ws = wb.active
    ws.title = 'SeasonalForecast'

    # Headers row 1
    headers = ['Quarter', 'Year 1', 'Year 2', 'Year 3', 'Avg', 'Overall Avg', 'Seasonal Index']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Historical data rows 2-5 (Q1-Q4)
    # Columns E, F, G (Avg, Overall Avg, Seasonal Index) are intentionally left empty
    history = [
        ('Q1', 1800000, 2100000, 2400000),
        ('Q2', 2200000, 2600000, 2900000),
        ('Q3', 2100000, 2500000, 2800000),
        ('Q4', 3200000, 3800000, 4300000),
    ]
    for r, (qtr, y1, y2, y3) in enumerate(history, 2):
        ws.cell(row=r, column=1, value=qtr)
        ws.cell(row=r, column=2, value=y1)
        ws.cell(row=r, column=3, value=y2)
        ws.cell(row=r, column=4, value=y3)
        # Columns E (5), F (6), G (7) left empty on purpose

    # Row 7: blank separator

    # Row 8: Next Year Forecast section header
    ws.cell(row=8, column=1, value='Next Year Forecast').font = Font(bold=True)
    ws.cell(row=8, column=2, value='Raw Trend Forecast').font = Font(bold=True)
    ws.cell(row=8, column=3, value='Adjusted Forecast').font = Font(bold=True)

    # Rows 9-12: forecast quarters with blank values (to be filled by agent)
    forecast_quarters = ['Q1', 'Q2', 'Q3', 'Q4']
    for r, qtr in enumerate(forecast_quarters, 9):
        ws.cell(row=r, column=1, value=qtr)
        # Raw Trend Forecast and Adjusted Forecast columns are intentionally empty

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
