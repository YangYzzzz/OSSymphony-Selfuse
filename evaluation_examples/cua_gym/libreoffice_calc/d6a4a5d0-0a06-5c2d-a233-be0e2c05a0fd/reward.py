"""
Reward Script: VLOOKUP with MATCH dynamic column lookup
Task ID: calc_lf_015
Domain: libreoffice_calc
Scoring:
  Component 1 (0.2): H2 contains a formula (not empty/static value)
  Component 2 (0.3): H2 formula uses VLOOKUP function
  Component 3 (0.2): H2 formula uses MATCH function
  Component 4 (0.3): H2 formula is functionally correct (proper ranges, lookup value, exact match)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_015'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Sales' sheet must exist
    if 'Sales' not in wb.sheetnames:
        print("FAIL: 'Sales' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Sales']

    # Get H2 value
    h2_val = ws['H2'].value

    # Component 1: H2 contains a formula (0.2 points)
    # In initial_env, H2 is None. In golden_env, H2 has a formula string starting with '='
    try:
        if h2_val is not None and isinstance(h2_val, str) and h2_val.startswith('='):
            print(f"PASS: Component 1 — H2 contains a formula: {h2_val[:50]}... (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 1 — H2 does not contain a formula, found: {h2_val!r}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: H2 formula uses VLOOKUP (0.3 points)
    # The task specifically asks for VLOOKUP with MATCH
    try:
        if isinstance(h2_val, str) and 'VLOOKUP' in h2_val.upper():
            print(f"PASS: Component 2 — H2 formula contains VLOOKUP (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — H2 formula does not contain VLOOKUP, found: {h2_val!r}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: H2 formula uses MATCH (0.2 points)
    # The task requires MATCH for dynamic column lookup
    try:
        if isinstance(h2_val, str) and 'MATCH' in h2_val.upper():
            print(f"PASS: Component 3 — H2 formula contains MATCH (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — H2 formula does not contain MATCH, found: {h2_val!r}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: H2 formula is functionally correct (0.3 points)
    # The formula should:
    #   - Look up "Widget A" (the lookup value)
    #   - Use a table range covering A1:E4 (or equivalent covering all data)
    #   - Use MATCH(G2, header_range, 0) for dynamic column index
    #   - Use exact match (last arg 0 or FALSE)
    # Expected: =VLOOKUP("Widget A",A1:E4,MATCH(G2,A1:E1,0),0)
    try:
        if isinstance(h2_val, str):
            formula_upper = h2_val.upper().replace(' ', '')

            checks_passed = 0
            total_sub_checks = 3

            # Sub-check 4a: Formula references "Widget A" as lookup value
            if 'WIDGETA' in formula_upper.replace('"', '').replace("'", "") or \
               '"WIDGET A"' in h2_val.upper().replace(' ', '').replace('\xa0', '') or \
               '"Widget A"' in h2_val or "'Widget A'" in h2_val:
                checks_passed += 1
                print("  Sub-check 4a: lookup value 'Widget A' found")
            else:
                print(f"  Sub-check 4a FAIL: 'Widget A' not found as lookup value")

            # Sub-check 4b: MATCH references G2 for the month selector
            if 'G2' in formula_upper:
                checks_passed += 1
                print("  Sub-check 4b: G2 reference found in formula")
            else:
                print(f"  Sub-check 4b FAIL: G2 reference not found")

            # Sub-check 4c: Uses exact match (0 or FALSE as last VLOOKUP arg)
            # The formula should end with ,0) or ,FALSE)
            if re.search(r',\s*(0|FALSE)\s*\)\s*$', h2_val, re.IGNORECASE):
                checks_passed += 1
                print("  Sub-check 4c: Exact match (0/FALSE) found")
            else:
                print(f"  Sub-check 4c FAIL: exact match flag not found at end of formula")

            if checks_passed == total_sub_checks:
                print(f"PASS: Component 4 — Formula is functionally correct (0.3 pts)")
                total_score += 0.3
            elif checks_passed > 0:
                partial = round(0.3 * checks_passed / total_sub_checks, 2)
                print(f"PARTIAL: Component 4 — {checks_passed}/{total_sub_checks} sub-checks passed ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 4 — No sub-checks passed")
        else:
            print(f"FAIL: Component 4 — H2 is not a formula string: {h2_val!r}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
