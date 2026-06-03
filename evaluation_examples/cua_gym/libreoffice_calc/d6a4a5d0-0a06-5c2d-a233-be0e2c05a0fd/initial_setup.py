"""
Initial Setup: VLOOKUP with MATCH dynamic column lookup
Task ID: calc_lf_015
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_015'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sales'

    # Headers
    ws['A1'] = 'Product'
    ws['B1'] = 'Jan'
    ws['C1'] = 'Feb'
    ws['D1'] = 'Mar'
    ws['E1'] = 'Apr'

    # Data rows
    data = [
        ['Widget A', 1200, 1350, 1500, 1425],
        ['Widget B', 980, 1100, 1050, 1200],
        ['Widget C', 2100, 2300, 2150, 2500],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Lookup area
    ws['G1'] = 'Month'
    ws['G2'] = 'Mar'
    ws['H1'] = 'Result'
    # H2 intentionally left empty - this is where the agent must enter the formula

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
