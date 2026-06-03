"""
Initial Setup: Pie chart with default starting angle (3 o'clock position)
Task ID: calc_chart_pie_start_angle_051
Domain: libreoffice_calc

Creates a spreadsheet with:
- Sheet 'CategorySales' with sales data by category
- A pie chart with firstSliceAng=90 (Electronics slice starting at 3 o'clock / right side)
"""

import openpyxl
from openpyxl.chart import PieChart, Reference

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_pie_start_angle_051'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: CategorySales ---
    ws = wb.active
    ws.title = 'CategorySales'

    # Headers
    ws['A1'] = 'Category'
    ws['B1'] = 'Sales'

    # Data - realistic sales figures by category
    data = [
        ('Electronics', 42000),
        ('Clothing', 18500),
        ('Home & Garden', 15200),
        ('Sports', 11800),
        ('Books', 8200),
    ]
    for r, (category, sales) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=category)
        ws.cell(row=r, column=2, value=sales)

    # Column widths for readability
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12

    # Create pie chart with default-like starting angle (3 o'clock = firstSliceAng=90 in OOXML)
    pie = PieChart()
    pie.title = 'Sales by Category'
    pie.style = 10

    # Data reference: Sales column including header
    data_ref = Reference(ws, min_col=2, min_row=1, max_row=6)
    # Category labels
    labels_ref = Reference(ws, min_col=1, min_row=2, max_row=6)

    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(labels_ref)

    # Set starting angle to 90 degrees (3 o'clock / right side - the "default" position)
    # In OOXML: firstSliceAng=0 means top (12 o'clock), firstSliceAng=90 means 3 o'clock
    pie.firstSliceAng = 90

    # Place chart on the sheet
    ws.add_chart(pie, 'D2')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheet: CategorySales')
    print('Data rows: 5 categories with sales figures')
    print('Chart: Pie chart with firstSliceAng=90 (3 o\'clock / right side)')


create_initial()
