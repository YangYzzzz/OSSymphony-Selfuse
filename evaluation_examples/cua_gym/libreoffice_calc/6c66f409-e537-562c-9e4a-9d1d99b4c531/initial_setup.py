"""
Initial Setup: Freelancer Time Tracking Spreadsheet
Task ID: calc_gen_freelancer_030
Domain: libreoffice_calc

Creates initial state:
- TimeLog sheet with 150 entries (Date, Client, Project, Hours populated; Billable/Rate/Amount/Invoiced empty)
- ClientRates sheet with 6 clients and hourly rates
- InvoiceSummary sheet (empty)
"""

import os
import openpyxl
from datetime import date, timedelta
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_freelancer_030'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

# Seed for reproducibility
random.seed(42)

# Client and rate data
CLIENTS = [
    'Acme Corp',
    'BlueSky Media',
    'ClearPath Tech',
    'DeltaForce Solutions',
    'EverGreen Designs',
    'FusionWave Studios',
]

RATES = {
    'Acme Corp': 150,
    'BlueSky Media': 100,
    'ClearPath Tech': 200,
    'DeltaForce Solutions': 125,
    'EverGreen Designs': 75,
    'FusionWave Studios': 175,
}

PROJECTS = {
    'Acme Corp': ['Website Redesign', 'SEO Audit', 'Content Strategy', 'CRM Integration'],
    'BlueSky Media': ['Social Campaign', 'Video Production', 'Brand Identity', 'Ad Copywriting'],
    'ClearPath Tech': ['API Development', 'Cloud Migration', 'Security Review', 'DevOps Setup'],
    'DeltaForce Solutions': ['Sales Dashboard', 'CRM Customization', 'Training Docs', 'Data Analysis'],
    'EverGreen Designs': ['Logo Design', 'Brochure Layout', 'UI Mockups', 'Illustration Pack'],
    'FusionWave Studios': ['Mobile App UI', 'Prototype Testing', 'Animation Assets', 'UX Research'],
}

def create_initial():
    wb = openpyxl.Workbook()

    # ---- Sheet 1: TimeLog ----
    ws_log = wb.active
    ws_log.title = 'TimeLog'

    headers = ['Date', 'Client', 'Project', 'Hours', 'Billable', 'Rate', 'Amount', 'Invoiced']
    for col, h in enumerate(headers, 1):
        ws_log.cell(row=1, column=col, value=h)

    # Generate 150 realistic time log entries
    # Spread across ~3 months ending at end of Feb 2026
    start_date = date(2025, 12, 1)
    end_date = date(2026, 2, 28)
    total_days = (end_date - start_date).days + 1

    entries = []
    for i in range(150):
        # Spread entries across the date range, skipping weekends sometimes
        day_offset = int(i * (total_days - 1) / 149)
        entry_date = start_date + timedelta(days=day_offset)
        # Occasionally skip a day (weekend or day off) - just shift by 1
        if entry_date.weekday() >= 5:  # Saturday=5, Sunday=6
            entry_date = entry_date + timedelta(days=(7 - entry_date.weekday()))

        client = CLIENTS[i % len(CLIENTS)]
        project = PROJECTS[client][i % len(PROJECTS[client])]
        # Hours: 0.5 to 8.0 in 0.5 increments
        hours_choices = [0.5, 1.0, 1.5, 2.0, 2.5, 3.0, 3.5, 4.0, 4.5, 5.0, 6.0, 7.0, 8.0]
        hours = random.choice(hours_choices)

        entries.append((entry_date, client, project, hours))

    # Sort by date
    entries.sort(key=lambda x: x[0])

    for r, (entry_date, client, project, hours) in enumerate(entries, 2):
        ws_log.cell(row=r, column=1, value=entry_date)  # Date
        ws_log.cell(row=r, column=2, value=client)       # Client
        ws_log.cell(row=r, column=3, value=project)      # Project
        ws_log.cell(row=r, column=4, value=hours)         # Hours
        # Columns E (Billable), F (Rate), G (Amount), H (Invoiced) — LEFT EMPTY intentionally
        # The agent must fill these in

    # Format date column
    for r in range(2, 152):
        ws_log.cell(row=r, column=1).number_format = 'yyyy-mm-dd'

    # ---- Sheet 2: ClientRates ----
    ws_rates = wb.create_sheet('ClientRates')

    ws_rates['A1'] = 'Client'
    ws_rates['B1'] = 'Hourly Rate'

    for i, client in enumerate(CLIENTS, 2):
        ws_rates.cell(row=i, column=1, value=client)
        ws_rates.cell(row=i, column=2, value=RATES[client])

    # ---- Sheet 3: InvoiceSummary ----
    ws_invoice = wb.create_sheet('InvoiceSummary')
    # Intentionally left empty - the agent must populate this

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  TimeLog: 150 rows of data (E, F, G, H columns empty)')
    print(f'  ClientRates: {len(CLIENTS)} clients with hourly rates')
    print(f'  InvoiceSummary: empty (to be filled by agent)')

create_initial()
