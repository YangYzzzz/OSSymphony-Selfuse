"""
Reward Script: Time Tracking and Billing Sheet for Freelancer
Task ID: calc_gen_freelancer_030
Domain: libreoffice_calc
Scoring:
  - Component 1: TimeLog F2:F151 has VLOOKUP rate formulas (0.25 pts)
  - Component 2: TimeLog E2:E151 has billable formulas (0.15 pts)
  - Component 3: TimeLog G2:G151 has amount formulas D*F (0.25 pts)
  - Component 4: TimeLog H2:H151 has Yes/No data validation (0.10 pts)
  - Component 5: InvoiceSummary has client SUMIFS summary rows (0.15 pts)
  - Component 6: InvoiceSummary has effective hourly rate calculation (0.10 pts)
Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_gen_freelancer_030'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Check required sheets exist
    required_sheets = ['TimeLog', 'ClientRates', 'InvoiceSummary']
    for sheet_name in required_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"CRITICAL: Required sheet '{sheet_name}' missing. Cannot verify task.")
            print("REWARD: 0.0")
            return 0.0

    ws_timelog = wb['TimeLog']
    ws_invoice = wb['InvoiceSummary']

    # Component 1: TimeLog F2:F151 has VLOOKUP rate formulas (0.25 pts)
    # These formulas should reference ClientRates to auto-fill hourly rates
    try:
        vlookup_count = 0
        sample_rows = list(range(2, 152))  # rows 2-151
        for row in sample_rows:
            cell_val = ws_timelog.cell(row=row, column=6).value  # Column F
            if cell_val is not None and isinstance(cell_val, str):
                upper_val = cell_val.upper()
                if 'VLOOKUP' in upper_val and 'CLIENTRATES' in upper_val:
                    vlookup_count += 1

        if vlookup_count >= 140:  # Allow some tolerance
            print(f"PASS: Component 1 — VLOOKUP rate formulas in F2:F151, found {vlookup_count}/150 rows (0.25 pts)")
            total_score += 0.25
        elif vlookup_count >= 70:
            print(f"PARTIAL: Component 1 — VLOOKUP rate formulas in F2:F151, found {vlookup_count}/150 rows (0.12 pts)")
            total_score += 0.12
        else:
            print(f"FAIL: Component 1 — Expected VLOOKUP formulas in F2:F151 referencing ClientRates, found {vlookup_count}/150")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: TimeLog E2:E151 has billable formulas (0.15 pts)
    # E column should have formulas that represent billable hours (=D2 or similar)
    try:
        billable_formula_count = 0
        for row in range(2, 152):  # rows 2-151
            cell_val = ws_timelog.cell(row=row, column=5).value  # Column E
            if cell_val is not None and isinstance(cell_val, str) and cell_val.startswith('='):
                billable_formula_count += 1

        if billable_formula_count >= 140:  # Allow some tolerance
            print(f"PASS: Component 2 — Billable formulas in E2:E151, found {billable_formula_count}/150 rows (0.15 pts)")
            total_score += 0.15
        elif billable_formula_count >= 70:
            print(f"PARTIAL: Component 2 — Billable formulas in E2:E151, found {billable_formula_count}/150 rows (0.07 pts)")
            total_score += 0.07
        else:
            print(f"FAIL: Component 2 — Expected formulas in E2:E151, found {billable_formula_count}/150")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: TimeLog G2:G151 has amount formulas D*F (0.25 pts)
    # G column should have amount = hours * rate formulas
    try:
        amount_formula_count = 0
        for row in range(2, 152):  # rows 2-151
            cell_val = ws_timelog.cell(row=row, column=7).value  # Column G
            if cell_val is not None and isinstance(cell_val, str):
                # Check for multiplication formulas involving D and F columns
                # e.g., =D2*F2, =E2*F2, etc.
                upper_val = cell_val.upper().replace(' ', '')
                if re.search(r'=[DEG]\d+\*[DEF]\d+|=[DEF]\d+\*[DEG]\d+', upper_val):
                    amount_formula_count += 1
                elif re.search(r'=D\d+\*F\d+|=F\d+\*D\d+', upper_val):
                    amount_formula_count += 1

        if amount_formula_count >= 140:  # Allow some tolerance
            print(f"PASS: Component 3 — Amount formulas in G2:G151, found {amount_formula_count}/150 rows (0.25 pts)")
            total_score += 0.25
        elif amount_formula_count >= 70:
            print(f"PARTIAL: Component 3 — Amount formulas in G2:G151, found {amount_formula_count}/150 rows (0.12 pts)")
            total_score += 0.12
        else:
            print(f"FAIL: Component 3 — Expected amount formulas in G2:G151 (like =D2*F2), found {amount_formula_count}/150")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: TimeLog H2:H151 has Yes/No data validation (0.10 pts)
    # H column should have dropdown validation for Invoiced status
    try:
        validations = list(ws_timelog.data_validations.dataValidation)
        invoice_dv_found = False
        for dv in validations:
            if dv.type == 'list' and dv.formula1 and 'Yes' in str(dv.formula1) and 'No' in str(dv.formula1):
                sqref_str = str(dv.sqref)
                if 'H' in sqref_str:
                    invoice_dv_found = True
                    print(f"PASS: Component 4 — Yes/No data validation in H column found (sqref: {sqref_str}) (0.10 pts)")
                    total_score += 0.10
                    break

        if not invoice_dv_found:
            # Check if any list validation exists in the sheet at all
            if len(validations) > 0:
                print(f"FAIL: Component 4 — Found {len(validations)} validation(s) but none is Yes/No in H column. Validations: {[(dv.type, dv.formula1, str(dv.sqref)) for dv in validations]}")
            else:
                print(f"FAIL: Component 4 — No data validation found in TimeLog sheet. Expected Yes/No dropdown for H2:H151")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: InvoiceSummary has client SUMIFS summary rows (0.15 pts)
    # The InvoiceSummary sheet should have per-client summaries with SUMIFS formulas
    try:
        invoice_rows = ws_invoice.max_row
        sumifs_count = 0

        if invoice_rows >= 2:
            for row in range(1, invoice_rows + 1):
                for col in range(1, 5):
                    cell_val = ws_invoice.cell(row=row, column=col).value
                    if cell_val is not None and isinstance(cell_val, str) and 'SUMIFS' in cell_val.upper():
                        sumifs_count += 1

        # Expect at least 6 clients * 2 columns (hours + amounts) = 12 SUMIFS formulas
        if sumifs_count >= 10:
            print(f"PASS: Component 5 — InvoiceSummary has client SUMIFS summaries, found {sumifs_count} SUMIFS formulas (0.15 pts)")
            total_score += 0.15
        elif sumifs_count >= 4:
            print(f"PARTIAL: Component 5 — InvoiceSummary has some SUMIFS formulas ({sumifs_count}), expected at least 12 (0.07 pts)")
            total_score += 0.07
        else:
            print(f"FAIL: Component 5 — Expected SUMIFS formulas in InvoiceSummary for per-client totals, found {sumifs_count}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: InvoiceSummary has effective hourly rate calculation (0.10 pts)
    # Should have a cell showing total revenue / total hours = effective rate
    try:
        effective_rate_found = False
        invoice_rows = ws_invoice.max_row

        # Look for a cell that contains text about "effective" rate or a formula dividing totals
        for row in range(1, invoice_rows + 1):
            for col in range(1, 5):
                cell_val = ws_invoice.cell(row=row, column=col).value
                if cell_val is not None:
                    if isinstance(cell_val, str):
                        if 'effective' in cell_val.lower() or 'effective hourly rate' in cell_val.lower():
                            effective_rate_found = True
                            print(f"PASS: Component 6 — Effective hourly rate label found at ({row},{col}) (0.10 pts)")
                            total_score += 0.10
                            break
                        # Check adjacent formula cells that do division
                        elif '/' in cell_val and cell_val.startswith('='):
                            # Check if this cell is labeled as effective rate
                            label_cell = ws_invoice.cell(row=row, column=col - 1).value
                            if label_cell and 'effective' in str(label_cell).lower():
                                effective_rate_found = True
                                print(f"PASS: Component 6 — Effective hourly rate formula found at ({row},{col}) (0.10 pts)")
                                total_score += 0.10
                                break
            if effective_rate_found:
                break

        if not effective_rate_found:
            # Check if there's IFERROR(x/y,0) pattern somewhere in InvoiceSummary
            for row in range(1, invoice_rows + 1):
                for col in range(1, 5):
                    cell_val = ws_invoice.cell(row=row, column=col).value
                    if cell_val is not None and isinstance(cell_val, str):
                        if 'IFERROR' in cell_val.upper() and '/' in cell_val:
                            effective_rate_found = True
                            print(f"PASS: Component 6 — Effective hourly rate IFERROR division formula found at ({row},{col}) (0.10 pts)")
                            total_score += 0.10
                            break
                if effective_rate_found:
                    break

        if not effective_rate_found:
            print(f"FAIL: Component 6 — No effective hourly rate calculation found in InvoiceSummary. Expected formula like =IFERROR(TotalRevenue/TotalHours, 0)")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
