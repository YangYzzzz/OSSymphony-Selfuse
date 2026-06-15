"""
Reward Script: Apply AutoFormat style then override header row fill to dark navy #003366
Task ID: calc_gg2_015
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): Header row A1:F1 fill is dark navy (#003366)
  Component 2 (0.35): Data rows 2-20 have alternating banding (solid fills, at least 2 distinct colors)
  Component 3 (0.15): Header row cells are bold (from AutoFormat)
  Component 4 (0.15): All header cells have fill AND data rows have fill (comprehensive formatting applied)
"""

import openpyxl
import os

WORKDIR = '/home/user'
TASK_ID = 'calc_gg2_015'


def persist_app_state(domain: str):
    """Save any unsaved LibreOffice changes before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check the 'Sales Report' sheet exists
    if 'Sales Report' not in wb.sheetnames:
        print("CRITICAL: 'Sales Report' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Sales Report']

    # Component 1: Header row A1:F1 fill is dark navy #003366 (0.35 points)
    # This is the manual override step - the key task-introduced change.
    # Initial state has NO fill on header, so this discriminates correctly.
    try:
        navy_count = 0
        target_rgb = "FF003366"
        for c in range(1, 7):
            cell = ws.cell(row=1, column=c)
            fill_type = cell.fill.fill_type
            fill_rgb = None
            try:
                fill_rgb = cell.fill.fgColor.rgb
            except:
                pass
            if fill_type == "solid" and fill_rgb == target_rgb:
                navy_count += 1

        if navy_count == 6:
            print(f"PASS: Component 1 — All 6 header cells have dark navy fill #{target_rgb} (0.35 pts)")
            total_score += 0.35
        elif navy_count >= 4:
            partial = round(0.35 * navy_count / 6, 2)
            print(f"PARTIAL: Component 1 — {navy_count}/6 header cells have navy fill ({partial} pts)")
            total_score += partial
        else:
            # Check if header has any dark fill close to navy (tolerance for slight color differences)
            dark_fill_count = 0
            for c in range(1, 7):
                cell = ws.cell(row=1, column=c)
                fill_rgb = None
                try:
                    fill_rgb = cell.fill.fgColor.rgb
                except:
                    pass
                if fill_rgb and cell.fill.fill_type == "solid":
                    # Parse the RGB and check if it's a dark navy-ish color
                    try:
                        r_val = int(fill_rgb[2:4], 16)
                        g_val = int(fill_rgb[4:6], 16)
                        b_val = int(fill_rgb[6:8], 16)
                        # Dark navy: low R, low-mid G, mid-high B, all relatively dark
                        if r_val <= 20 and g_val <= 80 and b_val >= 80 and b_val <= 150:
                            dark_fill_count += 1
                    except:
                        pass
            if dark_fill_count >= 4:
                partial = round(0.35 * dark_fill_count / 6, 2)
                print(f"PARTIAL: Component 1 — {dark_fill_count}/6 headers have dark navy-like fill ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 1 — Only {navy_count}/6 header cells have navy fill, {dark_fill_count} have dark fill")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Data rows 2-20 have alternating banding (0.35 points)
    # AutoFormat applies alternating row colors. Initial has NO fill at all.
    # We check that rows have solid fills and there are at least 2 distinct colors alternating.
    try:
        row_fills = []
        filled_rows = 0
        for r in range(2, 21):
            cell = ws.cell(row=r, column=1)
            fill_type = cell.fill.fill_type
            fill_rgb = None
            try:
                fill_rgb = cell.fill.fgColor.rgb
            except:
                pass
            if fill_type == "solid" and fill_rgb and fill_rgb != "00000000":
                filled_rows += 1
                row_fills.append(fill_rgb)
            else:
                row_fills.append(None)

        distinct_colors = set(f for f in row_fills if f is not None)

        if filled_rows >= 15 and len(distinct_colors) >= 2:
            # Check alternating pattern: consecutive rows should differ
            alternating_count = 0
            for i in range(len(row_fills) - 1):
                if row_fills[i] is not None and row_fills[i + 1] is not None and row_fills[i] != row_fills[i + 1]:
                    alternating_count += 1
            # With 19 data rows, perfect alternating = 18 transitions
            if alternating_count >= 14:
                print(f"PASS: Component 2 — {filled_rows}/19 data rows filled, {len(distinct_colors)} distinct colors, {alternating_count} alternations (0.35 pts)")
                total_score += 0.35
            else:
                partial = round(0.35 * 0.7, 2)  # banding present but not perfectly alternating
                print(f"PARTIAL: Component 2 — banding present but only {alternating_count} alternations ({partial} pts)")
                total_score += partial
        elif filled_rows >= 10 and len(distinct_colors) >= 2:
            partial = round(0.35 * 0.5, 2)
            print(f"PARTIAL: Component 2 — some banding: {filled_rows} filled rows, {len(distinct_colors)} colors ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — insufficient banding: {filled_rows} filled rows, {len(distinct_colors)} distinct colors")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Header row cells are bold (0.15 points)
    # AutoFormat typically makes headers bold. Initial state has NO bold.
    try:
        bold_count = 0
        for c in range(1, 7):
            cell = ws.cell(row=1, column=c)
            if cell.font.bold:
                bold_count += 1

        if bold_count == 6:
            print(f"PASS: Component 3 — All 6 header cells are bold (0.15 pts)")
            total_score += 0.15
        elif bold_count >= 4:
            partial = round(0.15 * bold_count / 6, 2)
            print(f"PARTIAL: Component 3 — {bold_count}/6 header cells are bold ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {bold_count}/6 header cells are bold")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Comprehensive formatting — all 6 header cells have solid fill
    # AND at least some data rows also have solid fill (0.15 points)
    # This ensures both steps (AutoFormat + override) were done.
    # Initial state has NO solid fills anywhere, so this scores 0 on initial.
    try:
        header_filled = 0
        for c in range(1, 7):
            cell = ws.cell(row=1, column=c)
            if cell.fill.fill_type == "solid":
                try:
                    rgb = cell.fill.fgColor.rgb
                    if rgb and rgb != "00000000":
                        header_filled += 1
                except:
                    pass

        data_filled = 0
        for r in range(2, 21):
            cell = ws.cell(row=r, column=1)
            if cell.fill.fill_type == "solid":
                try:
                    rgb = cell.fill.fgColor.rgb
                    if rgb and rgb != "00000000":
                        data_filled += 1
                except:
                    pass

        if header_filled == 6 and data_filled >= 15:
            print(f"PASS: Component 4 — All headers filled ({header_filled}/6) and data rows filled ({data_filled}/19) (0.15 pts)")
            total_score += 0.15
        elif header_filled >= 4 and data_filled >= 10:
            print(f"PARTIAL: Component 4 — Headers: {header_filled}/6, data rows: {data_filled}/19 (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 4 — Headers filled: {header_filled}/6, data rows filled: {data_filled}/19")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
