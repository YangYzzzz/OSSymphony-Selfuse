"""
Initial Setup: Apply AutoFormat style and override header fill color
Task ID: calc_gg2_015
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg2_015'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sales Report'

    # Headers in row 1
    headers = ['Order ID', 'Product', 'Region', 'Quantity', 'Unit Price', 'Total']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 19 rows of realistic sales data (rows 2-20)
    data = [
        ['ORD-2401', 'Wireless Keyboard', 'North America', 15, 49.99, 749.85],
        ['ORD-2402', 'USB-C Hub Adapter', 'Europe', 32, 34.50, 1104.00],
        ['ORD-2403', 'Noise-Canceling Headphones', 'Asia Pacific', 8, 189.00, 1512.00],
        ['ORD-2404', 'Ergonomic Mouse', 'North America', 45, 29.95, 1347.75],
        ['ORD-2405', 'Portable SSD 1TB', 'Europe', 12, 119.99, 1439.88],
        ['ORD-2406', 'Webcam HD 1080p', 'Latin America', 27, 59.00, 1593.00],
        ['ORD-2407', 'Laptop Stand Adjustable', 'Asia Pacific', 19, 42.50, 807.50],
        ['ORD-2408', 'Bluetooth Speaker', 'North America', 38, 75.00, 2850.00],
        ['ORD-2409', 'Monitor Arm Dual', 'Europe', 6, 149.99, 899.94],
        ['ORD-2410', 'Mechanical Keyboard', 'North America', 22, 89.95, 1978.90],
        ['ORD-2411', 'Desk Lamp LED', 'Asia Pacific', 55, 24.99, 1374.45],
        ['ORD-2412', 'Cable Management Kit', 'Latin America', 40, 15.50, 620.00],
        ['ORD-2413', 'Docking Station', 'Europe', 10, 199.00, 1990.00],
        ['ORD-2414', 'Wireless Charger Pad', 'North America', 63, 22.00, 1386.00],
        ['ORD-2415', 'Privacy Screen Filter', 'Asia Pacific', 17, 39.99, 679.83],
        ['ORD-2416', 'USB Flash Drive 128GB', 'Latin America', 90, 12.99, 1169.10],
        ['ORD-2417', 'Surge Protector Strip', 'North America', 35, 28.50, 997.50],
        ['ORD-2418', 'External DVD Drive', 'Europe', 14, 45.00, 630.00],
        ['ORD-2419', 'Trackball Mouse', 'Asia Pacific', 9, 64.95, 584.55],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12

    # No formatting applied - plain data only
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
