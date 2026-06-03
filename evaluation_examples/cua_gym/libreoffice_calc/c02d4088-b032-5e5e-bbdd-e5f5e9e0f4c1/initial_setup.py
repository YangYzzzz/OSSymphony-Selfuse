"""
Initial Setup: Create Order_Tracker spreadsheet with 59 orders, no conditional formatting.
Task ID: calc_gcv_018
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_018'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Order_Tracker"

    # --- Headers ---
    headers = [
        "Order ID", "Customer", "Product", "Quantity", "Unit Price",
        "Total", "Ship Date", "Carrier", "Tracking", "Status"
    ]
    header_font = Font(name="Calibri", size=11, bold=True, color="000000")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # --- Realistic data ---
    customers = [
        "Sarah Chen", "Marcus Johnson", "Emily Rodriguez", "David Kim",
        "Rachel Patel", "James O'Brien", "Yuki Tanaka", "Sofia Morales",
        "Liam Foster", "Aisha Khan", "Thomas Wright", "Mia Anderson",
        "Daniel Lopez", "Chloe Wang", "Michael Brown", "Natalie Scott",
        "Kevin Nguyen", "Isabella Garcia", "Robert Taylor", "Priya Sharma",
        "William Lee", "Olivia Martinez", "Ethan Davis", "Zoe Campbell",
        "Alexander Wilson", "Hannah Moore", "Benjamin Clark", "Lily Thompson",
        "Joshua Hall", "Emma Bennett"
    ]

    products = [
        "Wireless Mouse", "USB-C Hub", "Mechanical Keyboard", "Monitor Stand",
        "Laptop Sleeve", "Webcam HD", "Noise-Cancel Headphones", "Desk Lamp LED",
        "Portable SSD 1TB", "Ergonomic Chair", "Standing Desk Mat", "Cable Organizer",
        "Screen Protector", "Bluetooth Speaker", "Power Strip Surge", "USB Flash Drive 64GB",
        "Tablet Stand", "HDMI Cable 6ft", "Wireless Charger", "External DVD Drive",
        "Graphic Tablet", "Smart Pen", "Document Scanner", "Ring Light"
    ]

    carriers = ["FedEx", "UPS", "USPS", "DHL", "Amazon Logistics"]

    statuses = ["Shipped", "Processing", "Cancelled", "Returned"]
    # Distribute statuses: ~50% Shipped, ~25% Processing, ~15% Cancelled, ~10% Returned
    status_weights = [50, 25, 15, 10]

    random.seed(42)  # Reproducible

    for row_idx in range(2, 61):  # rows 2 to 60 = 59 orders
        order_num = row_idx - 1
        order_id = f"ORD-2025-{order_num:04d}"
        customer = random.choice(customers)
        product = random.choice(products)
        quantity = random.randint(1, 20)
        unit_price = round(random.uniform(9.99, 499.99), 2)
        total = round(quantity * unit_price, 2)

        # Ship dates between 2025-01-05 and 2025-03-28
        month = random.randint(1, 3)
        day = random.randint(1, 28)
        ship_date = f"2025-{month:02d}-{day:02d}"

        carrier = random.choice(carriers)
        tracking = f"{carrier[:2].upper()}{random.randint(100000000, 999999999)}"
        status = random.choices(statuses, weights=status_weights, k=1)[0]

        row_data = [order_id, customer, product, quantity, unit_price,
                    total, ship_date, carrier, tracking, status]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = Font(name="Calibri", size=11, color="000000")

    # --- Column widths ---
    col_widths = {
        "A": 16, "B": 20, "C": 26, "D": 10, "E": 12,
        "F": 12, "G": 14, "H": 18, "I": 16, "J": 12
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # No conditional formatting in initial state
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
