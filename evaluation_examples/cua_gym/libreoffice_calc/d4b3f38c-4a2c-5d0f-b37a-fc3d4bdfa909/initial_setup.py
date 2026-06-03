"""
Initial Setup: HR Job Level Framework spreadsheet (no formatting)
Task ID: calc_hr_job_level_framework_055
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_job_level_framework_055'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Level Framework ---
    ws = wb.active
    ws.title = 'Level Framework'

    # Row 1: unmerged, unformatted title
    ws['A1'] = 'Job Level Framework 2024'

    # Row 2: column headers (no formatting)
    headers = ['Level', 'Title', 'Band Min', 'Band Max', 'Equity Range', 'Tier']
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)

    # Rows 3-12: job level data (realistic content, no formatting)
    # Engineering track (L1-L5) + PM track (P1-P3) + Design track (D1-D2)
    data = [
        # Level, Title,               Band Min, Band Max, Equity Range,    Tier
        ['L1', 'Junior Software Engineer',   65000,  80000,  '0.05% - 0.10%', 'Junior'],
        ['L2', 'Software Engineer',          80000, 100000,  '0.10% - 0.20%', 'Junior'],
        ['L3', 'Senior Software Engineer',  100000, 130000,  '0.20% - 0.40%', 'Mid'],
        ['L4', 'Staff Engineer',            130000, 160000,  '0.40% - 0.70%', 'Senior'],
        ['L5', 'Principal Engineer',        160000, 200000,  '0.70% - 1.20%', 'Senior'],
        ['P1', 'Associate Product Manager',  70000,  90000,  '0.05% - 0.10%', 'Junior'],
        ['P2', 'Product Manager',            95000, 125000,  '0.15% - 0.30%', 'Mid'],
        ['P3', 'Senior Product Manager',    130000, 165000,  '0.35% - 0.65%', 'Senior'],
        ['D1', 'UX Designer',                75000,  95000,  '0.05% - 0.15%', 'Junior'],
        ['D2', 'Senior UX Designer',        100000, 130000,  '0.20% - 0.40%', 'Mid'],
    ]

    for r, row_data in enumerate(data, 3):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths for readability (not a formatting task requirement)
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
