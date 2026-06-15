"""
Reward Script: Create job level framework table with salary bands, borders, merged header,
               and conditional formatting for seniority tiers.
Task ID: calc_hr_job_level_framework_055
Domain: libreoffice_calc
Scoring:
  Component 1: A1:F1 merged with bold/16pt/centered styling          (0.25 pts)
  Component 2: Row 2 headers bold + gray background (#D9D9D9)        (0.20 pts)
  Component 3: Thin borders on all cells A2:F12                      (0.20 pts)
  Component 4: Conditional formatting rules on A3:F12 (4 tier rules) (0.20 pts)
  Component 5: Number format $#,##0 on C3:D12                        (0.15 pts)
  Total: 1.0
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_job_level_framework_055'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sheet 'Level Framework' must exist
    if 'Level Framework' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Level Framework' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Level Framework']

    # Component 1: A1:F1 merged with bold, 16pt, centered styling (0.25 points)
    # Task requires: merge A1:F1, bold, 16pt font, centered alignment
    # This FAILS on initial (no merge, no bold, no 16pt, no centered)
    # and PASSES on golden (all applied)
    try:
        # Check if A1:F1 appears in merged ranges
        a1_range_merged = any(str(rng) == 'A1:F1' for rng in ws.merged_cells.ranges)
        b1_merged = isinstance(ws.cell(row=1, column=2), MergedCell)
        cells_merged = a1_range_merged or b1_merged

        if cells_merged:
            # Base merge score
            merge_score = 0.10
            print("PASS: Component 1 base — A1:F1 is merged (+0.10 pts)")

            # Bold check
            a1 = ws.cell(row=1, column=1)
            if a1.font.bold is True:
                merge_score += 0.05
                print("PASS: Component 1a — A1 is bold (+0.05 pts)")
            else:
                print(f"FAIL: Component 1a — A1 is not bold (bold={a1.font.bold})")

            # 16pt font check
            if a1.font.size is not None and abs(a1.font.size - 16) < 0.5:
                merge_score += 0.05
                print(f"PASS: Component 1b — A1 font size is 16pt (size={a1.font.size}) (+0.05 pts)")
            else:
                print(f"FAIL: Component 1b — A1 font size is not 16pt (size={a1.font.size})")

            # Centered alignment check
            if a1.alignment.horizontal in ('center', 'centerContinuous'):
                merge_score += 0.05
                print(f"PASS: Component 1c — A1 alignment is centered (+0.05 pts)")
            else:
                print(f"FAIL: Component 1c — A1 alignment is not centered (alignment={a1.alignment.horizontal})")

            total_score += merge_score
            print(f"Component 1 subtotal: {merge_score}/0.25")
        else:
            print("FAIL: Component 1 — A1:F1 is not merged (0.0 pts)")

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Row 2 headers bold with gray background #D9D9D9 (0.20 points)
    # Task requires: all 6 headers in row 2 are bold and have #D9D9D9 background
    # This FAILS on initial (no bold, no background) and PASSES on golden
    try:
        bold_count = 0
        gray_count = 0
        expected_bg_argb = 'FFD9D9D9'

        for col in range(1, 7):
            cell = ws.cell(row=2, column=col)
            if cell.font.bold:
                bold_count += 1
            try:
                bg = cell.fill.fgColor.rgb
                if bg.upper() == expected_bg_argb:
                    gray_count += 1
            except Exception:
                pass

        if bold_count == 6:
            total_score += 0.10
            print(f"PASS: Component 2a — All 6 row 2 headers are bold (+0.10 pts)")
        else:
            print(f"FAIL: Component 2a — Only {bold_count}/6 headers are bold")

        if gray_count == 6:
            total_score += 0.10
            print(f"PASS: Component 2b — All 6 headers have gray background #D9D9D9 (+0.10 pts)")
        else:
            print(f"FAIL: Component 2b — Only {gray_count}/6 headers have correct gray background")

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Thin borders applied to all cells in A2:F12 (0.20 points)
    # Task requires: borders (thin, all sides) on A2:F12
    # This FAILS on initial (no borders) and PASSES on golden
    try:
        border_pass = 0
        border_total = 0

        for row in range(2, 13):  # rows 2-12
            for col in range(1, 7):  # cols A-F
                cell = ws.cell(row=row, column=col)
                border_total += 1
                if (cell.border.left.style == 'thin' and
                        cell.border.right.style == 'thin' and
                        cell.border.top.style == 'thin' and
                        cell.border.bottom.style == 'thin'):
                    border_pass += 1

        if border_pass == border_total:
            total_score += 0.20
            print(f"PASS: Component 3 — Thin borders on all {border_total} cells in A2:F12 (+0.20 pts)")
        elif border_pass >= int(border_total * 0.9):
            total_score += 0.10
            print(f"PARTIAL: Component 3 — {border_pass}/{border_total} cells have thin borders (+0.10 pts)")
        else:
            print(f"FAIL: Component 3 — Only {border_pass}/{border_total} cells have thin borders")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Conditional formatting rules on A3:F12 for tiers (0.20 points)
    # Task requires 4 rules: Junior=DEEAF1, Mid=E2EFDA, Senior=FFF2CC, Principal/Staff=FCE4D6
    # This FAILS on initial (no CF) and PASSES on golden
    try:
        rules_list = []
        cf_range_found = False

        for cf_range, rules in ws.conditional_formatting._cf_rules.items():
            range_str = str(cf_range)
            if 'A3' in range_str and 'F12' in range_str:
                cf_range_found = True
                for rule in rules:
                    try:
                        fill_color = rule.dxf.fill.fgColor.rgb.upper() if (rule.dxf and rule.dxf.fill) else None
                        formula = rule.formula[0] if rule.formula else ''
                        rules_list.append((formula, fill_color))
                    except Exception:
                        pass

        if not cf_range_found:
            print("FAIL: Component 4 — No conditional formatting found on A3:F12 range")
        else:
            # Check each tier color rule
            junior_ok = any('Junior' in f and c == 'FFDEEAF1' for f, c in rules_list)
            mid_ok = any('Mid' in f and c == 'FFE2EFDA' for f, c in rules_list)
            senior_ok = any('Senior' in f and c == 'FFFFF2CC' for f, c in rules_list
                            if 'Principal' not in f and 'Staff' not in f)
            principal_ok = any(('Principal' in f or 'Staff' in f) and c == 'FFFCE4D6' for f, c in rules_list)

            rules_matched = sum([junior_ok, mid_ok, senior_ok, principal_ok])

            if junior_ok:
                print("PASS: Component 4a — Junior tier rule (#DEEAF1) found")
            else:
                print("FAIL: Component 4a — Junior tier rule (#DEEAF1) not found")

            if mid_ok:
                print("PASS: Component 4b — Mid tier rule (#E2EFDA) found")
            else:
                print("FAIL: Component 4b — Mid tier rule (#E2EFDA) not found")

            if senior_ok:
                print("PASS: Component 4c — Senior tier rule (#FFF2CC) found")
            else:
                print("FAIL: Component 4c — Senior tier rule (#FFF2CC) not found")

            if principal_ok:
                print("PASS: Component 4d — Principal/Staff tier rule (#FCE4D6) found")
            else:
                print("FAIL: Component 4d — Principal/Staff tier rule (#FCE4D6) not found")

            cf_score = round(0.05 * rules_matched, 2)
            total_score += cf_score
            print(f"Component 4 subtotal: {cf_score}/0.20 ({rules_matched}/4 rules matched)")

    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Number format $#,##0 applied to C3:D12 (0.15 points)
    # Task requires: C3:D12 formatted as $#,##0
    # This FAILS on initial (General format) and PASSES on golden
    try:
        fmt_pass = 0
        fmt_total = 0
        expected_fmt = '$#,##0'

        for row in range(3, 13):
            for col in [3, 4]:  # columns C and D
                cell = ws.cell(row=row, column=col)
                fmt_total += 1
                if cell.number_format == expected_fmt:
                    fmt_pass += 1

        if fmt_pass == fmt_total:
            total_score += 0.15
            print(f"PASS: Component 5 — All {fmt_total} cells in C3:D12 have $#,##0 format (+0.15 pts)")
        elif fmt_pass > 0:
            partial = round(0.15 * fmt_pass / fmt_total, 2)
            total_score += partial
            print(f"PARTIAL: Component 5 — {fmt_pass}/{fmt_total} cells have $#,##0 format (+{partial} pts)")
        else:
            print(f"FAIL: Component 5 — No cells in C3:D12 have $#,##0 format")

    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
