"""
Reward Script: Paste Special > Paste Unformatted Text from web source
Task ID: calc_gsi_081
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): 5 new data rows (12-16) with correct product values
  Component 2 (0.35): Pasted rows have NO external formatting (no fills, no bold, clean font)
  Component 3 (0.30): Number formats on pasted rows match existing data pattern
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_081'

# Expected new data from the web HTML table (rows 12-16)
EXPECTED_NEW_ROWS = [
    {'row': 12, 'product': 'Mechanical Keyboard RGB', 'category': 'Electronics', 'units': 198, 'revenue': 15840.00, 'date': '2025-10-03'},
    {'row': 13, 'product': 'Noise-Cancel Headphones', 'category': 'Electronics', 'units': 87, 'revenue': 17399.13, 'date': '2025-10-08'},
    {'row': 14, 'product': 'Adjustable Desk Riser', 'category': 'Furniture', 'units': 45, 'revenue': 6749.55, 'date': '2025-10-15'},
    {'row': 15, 'product': 'USB Docking Station', 'category': 'Accessories', 'units': 156, 'revenue': 12479.44, 'date': '2025-10-20'},
    {'row': 16, 'product': 'Portable Charger 20000mAh', 'category': 'Electronics', 'units': 289, 'revenue': 8670.00, 'date': '2025-10-25'},
]


def _check_has_fill(cell):
    """Return whether a cell has a non-trivial background fill."""
    try:
        if cell.fill.patternType and cell.fill.patternType != 'none':
            fill_rgb = cell.fill.fgColor.rgb if cell.fill.fgColor else None
            if fill_rgb and fill_rgb not in ('00000000', '0'):
                return True  # non-trivial fill detected
    except Exception:
        pass
    return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Sales Report' sheet must exist
    if 'Sales Report' not in wb.sheetnames:
        print("FAIL: 'Sales Report' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Sales Report']

    # Check if we have enough rows -- if only 11 rows (initial state), no new data was pasted
    if ws.max_row < 12:
        print(f"FAIL: Sheet has only {ws.max_row} rows -- no new data pasted (expected at least 16)")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: New data rows exist with correct values (0.35 points)
    # Each of the 5 rows contributes 0.07 points
    try:
        comp1_score = 0.0
        rows_matched = 0
        for expected in EXPECTED_NEW_ROWS:
            r = expected['row']
            product_val = ws.cell(row=r, column=1).value
            category_val = ws.cell(row=r, column=2).value
            units_val = ws.cell(row=r, column=3).value
            revenue_val = ws.cell(row=r, column=4).value
            date_val = ws.cell(row=r, column=5).value

            # Check product name
            product_match = (str(product_val).strip() == expected['product']) if product_val else False
            # Check category
            category_match = (str(category_val).strip() == expected['category']) if category_val else False
            # Check units (may be int or string)
            units_match = False
            if units_val is not None:
                try:
                    units_match = int(float(str(units_val).replace(',', ''))) == expected['units']
                except (ValueError, TypeError):
                    pass
            # Check revenue (with tolerance)
            revenue_match = False
            if revenue_val is not None:
                try:
                    rev_num = float(str(revenue_val).replace('$', '').replace(',', ''))
                    revenue_match = abs(rev_num - expected['revenue']) < 0.02
                except (ValueError, TypeError):
                    pass
            # Check date
            date_match = False
            if date_val is not None:
                date_str = str(date_val).strip()
                # Could be datetime object or string
                if hasattr(date_val, 'strftime'):
                    date_str = date_val.strftime('%Y-%m-%d')
                date_match = date_str == expected['date']

            all_match = product_match and category_match and units_match and revenue_match and date_match
            if all_match:
                rows_matched += 1
                comp1_score += 0.07
                print(f"PASS: Row {r} data correct — {expected['product']}")
            else:
                print(f"FAIL: Row {r} data mismatch — product={product_val}(exp={expected['product']}), "
                      f"cat={category_val}(exp={expected['category']}), units={units_val}(exp={expected['units']}), "
                      f"rev={revenue_val}(exp={expected['revenue']}), date={date_val}(exp={expected['date']})")

        if rows_matched == 5:
            comp1_score = 0.35  # exact total
        print(f"Component 1 total: {comp1_score:.2f}/0.35 ({rows_matched}/5 rows matched)")
        total_score += comp1_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Pasted rows have NO external formatting (0.35 points)
    # The web HTML has Comic Sans, bold categories, colored text, background fills.
    # After Paste Special > Unformatted Text, none of these should be present.
    # We check rows 12-16 for absence of external styling.
    try:
        comp2_score = 0.0
        clean_rows = 0
        for r in range(12, 17):
            row_issues = 0
            for c in range(1, 6):
                cell = ws.cell(row=r, column=c)
                if cell.value is None:
                    continue

                # Check: no background fill on data cells
                cell_has_fill = _check_has_fill(cell)

                if cell_has_fill:
                    print(f"FAIL: Cell {cell.coordinate} has background fill (fill={cell.fill.fgColor.rgb})")
                    row_issues += 1

                # Check: data cells should NOT be bold (the web source has bold on category and revenue)
                if cell.font.bold:
                    print(f"FAIL: Cell {cell.coordinate} is bold (web formatting leaked)")
                    row_issues += 1

                # Check: font should NOT be Comic Sans MS (the web source uses Comic Sans)
                if cell.font.name and 'comic' in cell.font.name.lower():
                    print(f"FAIL: Cell {cell.coordinate} has Comic Sans font (web formatting leaked)")
                    row_issues += 1

                # Check: font color should not be the web's special colors (#8e44ad purple, #27ae60 green)
                try:
                    if cell.font.color and cell.font.color.rgb:
                        fc = cell.font.color.rgb.upper()
                        # Web source colors: purple #8e44ad, green #27ae60, dark #2c3e50
                        web_colors = ['008E44AD', 'FF8E44AD', '0027AE60', 'FF27AE60']
                        if fc in web_colors:
                            print(f"FAIL: Cell {cell.coordinate} has web-source font color ({fc})")
                            row_issues += 1
                except:
                    pass

            if row_issues == 0:
                clean_rows += 1
                print(f"PASS: Row {r} has clean formatting (no web styles)")

        # Each clean row is worth 0.07
        comp2_score = (clean_rows / 5) * 0.35
        if clean_rows == 5:
            comp2_score = 0.35
        print(f"Component 2 total: {comp2_score:.2f}/0.35 ({clean_rows}/5 rows clean)")
        total_score += comp2_score
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Number formats on pasted rows match existing data pattern (0.30 points)
    # Existing rows use: $#,##0.00 for Revenue (col D), yyyy-mm-dd for dates (col E), General for units (col C)
    try:
        comp3_score = 0.0
        format_checks_passed = 0
        total_format_checks = 0

        for r in range(12, 17):
            # Check Revenue column (D) has currency format
            rev_cell = ws.cell(row=r, column=4)
            total_format_checks += 1
            if rev_cell.value is not None:
                nf = rev_cell.number_format
                # Should be a currency-like format or the value should be numeric
                if '$' in str(nf) or isinstance(rev_cell.value, (int, float)):
                    format_checks_passed += 1
                    print(f"PASS: D{r} has correct format ({nf}, value type={type(rev_cell.value).__name__})")
                else:
                    print(f"FAIL: D{r} expected numeric/currency format, got '{nf}', value={rev_cell.value}")
            else:
                print(f"FAIL: D{r} is empty")

            # Check Units column (C) is numeric
            units_cell = ws.cell(row=r, column=3)
            total_format_checks += 1
            if units_cell.value is not None and isinstance(units_cell.value, (int, float)):
                format_checks_passed += 1
                print(f"PASS: C{r} is numeric ({units_cell.value})")
            else:
                print(f"FAIL: C{r} expected numeric value, got {type(units_cell.value).__name__}: {units_cell.value}")

            # Check Date column (E) has date format or is a proper date string
            date_cell = ws.cell(row=r, column=5)
            total_format_checks += 1
            if date_cell.value is not None:
                date_str = str(date_cell.value)
                nf = date_cell.number_format
                # Accept date format or proper date string
                has_date_format = ('yy' in str(nf).lower() or 'mm' in str(nf).lower() or 'dd' in str(nf).lower())
                is_date_string = len(date_str) >= 10 and '-' in date_str[:10]
                has_datetime = hasattr(date_cell.value, 'strftime')
                if has_date_format or is_date_string or has_datetime:
                    format_checks_passed += 1
                    print(f"PASS: E{r} has date format (nf={nf}, val={date_cell.value})")
                else:
                    print(f"FAIL: E{r} expected date format, got nf='{nf}', val={date_cell.value}")
            else:
                print(f"FAIL: E{r} is empty")

        if total_format_checks > 0:
            comp3_score = (format_checks_passed / total_format_checks) * 0.30
        print(f"Component 3 total: {comp3_score:.2f}/0.30 ({format_checks_passed}/{total_format_checks} format checks)")
        total_score += comp3_score
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved LibreOffice state before verification
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(1.0)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
persist_app_state()

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
