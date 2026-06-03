"""
Initial Setup: Paste Special > Paste Unformatted Text
Task ID: calc_gsi_081
Domain: libreoffice_calc

Creates a destination spreadsheet with clean, consistent styling and existing
data. Also creates an HTML file simulating "web browser table data" that the
user will copy and paste into the spreadsheet using Paste Special.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_081'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'
HTML_SOURCE = f'{WORKDIR}/web_table_data.html'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Destination Sheet: Sales Report ---
    ws = wb.active
    ws.title = 'Sales Report'

    # Define consistent destination styling
    header_font = Font(name='Liberation Sans', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    data_font = Font(name='Liberation Sans', size=10, color='000000')
    data_align = Alignment(horizontal='left', vertical='center')
    currency_format = '$#,##0.00'
    date_format = 'yyyy-mm-dd'
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    # Headers
    headers = ['Product', 'Category', 'Units Sold', 'Revenue', 'Sale Date']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Existing data rows (consistent formatting)
    existing_data = [
        ['Wireless Keyboard', 'Electronics', 145, 4347.55, '2025-08-12'],
        ['Ergonomic Mouse', 'Electronics', 203, 5068.97, '2025-08-15'],
        ['USB-C Hub', 'Accessories', 87, 3478.13, '2025-08-18'],
        ['Monitor Stand', 'Furniture', 62, 2479.38, '2025-08-22'],
        ['Webcam HD Pro', 'Electronics', 178, 8899.22, '2025-09-01'],
        ['Desk Lamp LED', 'Furniture', 95, 2374.05, '2025-09-05'],
        ['Laptop Sleeve 15"', 'Accessories', 312, 4679.88, '2025-09-10'],
        ['Bluetooth Speaker', 'Electronics', 134, 5359.66, '2025-09-14'],
        ['Cable Organizer', 'Accessories', 256, 2047.44, '2025-09-18'],
        ['Standing Desk Mat', 'Furniture', 73, 2554.27, '2025-09-22'],
    ]

    for r, row_data in enumerate(existing_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if c == 4:  # Revenue column
                cell.number_format = currency_format
            elif c == 5:  # Date column
                cell.number_format = date_format

    # Set column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    # --- Second Sheet: Summary (to add complexity) ---
    ws2 = wb.create_sheet('Summary')
    ws2['A1'] = 'Total Products'
    ws2['B1'] = 10
    ws2['A2'] = 'Total Revenue'
    ws2['B2'] = '=SUM(\'Sales Report\'!D2:D11)'
    ws2['B2'].number_format = currency_format
    ws2['A3'] = 'Report Generated'
    ws2['B3'] = '2025-09-25'
    for row in ws2.iter_rows(min_row=1, max_row=3, min_col=1, max_col=2):
        for cell in row:
            cell.font = Font(name='Liberation Sans', size=10)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Create HTML file simulating web browser table data with heavy formatting
    html_content = """<!DOCTYPE html>
<html>
<head><title>Q4 Product Sales - WebStore Analytics</title></head>
<body style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f0f0f0;">
<h2 style="color: #e74c3c; font-family: Georgia, serif;">Q4 Product Sales Data</h2>
<table style="border-collapse: collapse; background: #fff; box-shadow: 0 2px 8px rgba(0,0,0,0.15);">
<tr style="background: linear-gradient(135deg, #e74c3c, #c0392b); color: white; font-family: Georgia, serif; font-size: 14px;">
  <th style="padding: 12px; border: 2px solid #c0392b;">Product</th>
  <th style="padding: 12px; border: 2px solid #c0392b;">Category</th>
  <th style="padding: 12px; border: 2px solid #c0392b;">Units Sold</th>
  <th style="padding: 12px; border: 2px solid #c0392b;">Revenue</th>
  <th style="padding: 12px; border: 2px solid #c0392b;">Sale Date</th>
</tr>
<tr style="background: #fdf2f2; font-family: 'Comic Sans MS', cursive; color: #2c3e50;">
  <td style="padding: 10px; border: 1px solid #e74c3c;">Mechanical Keyboard RGB</td>
  <td style="padding: 10px; border: 1px solid #e74c3c; color: #8e44ad; font-weight: bold;">Electronics</td>
  <td style="padding: 10px; border: 1px solid #e74c3c; text-align: right;">198</td>
  <td style="padding: 10px; border: 1px solid #e74c3c; text-align: right; color: #27ae60; font-weight: bold;">$15,840.00</td>
  <td style="padding: 10px; border: 1px solid #e74c3c;">2025-10-03</td>
</tr>
<tr style="background: #ffffff; font-family: 'Comic Sans MS', cursive; color: #2c3e50;">
  <td style="padding: 10px; border: 1px solid #e74c3c;">Noise-Cancel Headphones</td>
  <td style="padding: 10px; border: 1px solid #e74c3c; color: #8e44ad; font-weight: bold;">Electronics</td>
  <td style="padding: 10px; border: 1px solid #e74c3c; text-align: right;">87</td>
  <td style="padding: 10px; border: 1px solid #e74c3c; text-align: right; color: #27ae60; font-weight: bold;">$17,399.13</td>
  <td style="padding: 10px; border: 1px solid #e74c3c;">2025-10-08</td>
</tr>
<tr style="background: #fdf2f2; font-family: 'Comic Sans MS', cursive; color: #2c3e50;">
  <td style="padding: 10px; border: 1px solid #e74c3c;">Adjustable Desk Riser</td>
  <td style="padding: 10px; border: 1px solid #e74c3c; color: #8e44ad; font-weight: bold;">Furniture</td>
  <td style="padding: 10px; border: 1px solid #e74c3c; text-align: right;">45</td>
  <td style="padding: 10px; border: 1px solid #e74c3c; text-align: right; color: #27ae60; font-weight: bold;">$6,749.55</td>
  <td style="padding: 10px; border: 1px solid #e74c3c;">2025-10-15</td>
</tr>
<tr style="background: #ffffff; font-family: 'Comic Sans MS', cursive; color: #2c3e50;">
  <td style="padding: 10px; border: 1px solid #e74c3c;">USB Docking Station</td>
  <td style="padding: 10px; border: 1px solid #e74c3c; color: #8e44ad; font-weight: bold;">Accessories</td>
  <td style="padding: 10px; border: 1px solid #e74c3c; text-align: right;">156</td>
  <td style="padding: 10px; border: 1px solid #e74c3c; text-align: right; color: #27ae60; font-weight: bold;">$12,479.44</td>
  <td style="padding: 10px; border: 1px solid #e74c3c;">2025-10-20</td>
</tr>
<tr style="background: #fdf2f2; font-family: 'Comic Sans MS', cursive; color: #2c3e50;">
  <td style="padding: 10px; border: 1px solid #e74c3c;">Portable Charger 20000mAh</td>
  <td style="padding: 10px; border: 1px solid #e74c3c; color: #8e44ad; font-weight: bold;">Electronics</td>
  <td style="padding: 10px; border: 1px solid #e74c3c; text-align: right;">289</td>
  <td style="padding: 10px; border: 1px solid #e74c3c; text-align: right; color: #27ae60; font-weight: bold;">$8,670.00</td>
  <td style="padding: 10px; border: 1px solid #e74c3c;">2025-10-25</td>
</tr>
</table>
<p style="color: #999; font-size: 11px; font-style: italic;">Source: WebStore Analytics Dashboard - Generated 2025-10-28</p>
</body>
</html>"""

    with open(HTML_SOURCE, 'w') as f:
        f.write(html_content)
    print(f'HTML source file created: {HTML_SOURCE}')

    # GUI-ready startup: open the spreadsheet and the HTML file in browser
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    launch_gui(f'google-chrome "{HTML_SOURCE}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc and Chrome with DISPLAY=:0')


create_initial()
