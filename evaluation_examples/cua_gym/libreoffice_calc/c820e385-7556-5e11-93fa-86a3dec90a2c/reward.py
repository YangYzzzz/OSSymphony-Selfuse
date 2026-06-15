"""
Reward Script: Freeze top two rows in inventory.xlsx Stock sheet
Task ID: calc_gg5_002
Domain: libreoffice_calc
Scoring:
  Component 1 (0.6): freeze_panes is set to 'A3' (freezes rows 1-2)
  Component 2 (0.2): Sheet 'Stock' exists and has expected header row intact
  Component 3 (0.2): Data integrity — row count and structure preserved
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_002'


def persist_app_state(domain: str):
    """Save any unsaved LibreOffice edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify that rows 1-2 are frozen in the Stock sheet.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Stock' sheet must exist
    if 'Stock' not in wb.sheetnames:
        print(f"CRITICAL: 'Stock' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Stock']

    # Component 1: freeze_panes is set to 'A3' (0.6 points)
    # This is THE core task requirement — freezing top two rows means cursor at A3.
    # This FAILS on initial (None) and PASSES on golden ('A3').
    try:
        fp = ws.freeze_panes
        if fp == 'A3':
            print(f"PASS: Component 1 — freeze_panes is 'A3' (rows 1-2 frozen) (0.6 pts)")
            total_score += 0.6
        elif fp is not None:
            # Partial credit: something is frozen but not exactly right
            # e.g. 'A2' freezes only row 1, 'B3' also freezes column A
            fp_str = str(fp)
            # If row part is 3 (rows 1-2 frozen) but column is not A, give partial
            import re
            match = re.match(r'([A-Z]+)(\d+)', fp_str)
            col_letter = match.group(1) if match else ''
            row_num = int(match.group(2)) if match else 0
            if row_num == 3:
                # Rows 1-2 are frozen, but extra columns may also be frozen
                print(f"PARTIAL: Component 1 — freeze_panes is '{fp}' (rows 1-2 frozen but extra column freeze) (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 1 — freeze_panes is '{fp}', expected 'A3'")
        else:
            print(f"FAIL: Component 1 — freeze_panes is None (no freeze)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Header row integrity anchored to freeze (0.2 points)
    # Verify that the frozen rows still contain the expected headers.
    # This component checks that freeze + header integrity are both present.
    # On initial_env: freeze_panes is None, so this component gates on freeze first.
    try:
        fp = ws.freeze_panes
        if fp is not None:
            # Parse row number from freeze_panes string (e.g. 'A3' -> 3)
            import re
            match = re.match(r'[A-Z]+(\d+)', str(fp))
            row_num = int(match.group(1)) if match else 0
            if row_num >= 3:
                # Rows are frozen; verify headers are intact
                expected_headers = ['Item', 'SKU', 'Quantity', 'Unit', 'Price', 'Reorder Level']
                actual_headers = [ws.cell(row=1, column=c).value for c in range(1, 7)]
                if actual_headers == expected_headers:
                    print(f"PASS: Component 2 — Headers intact in frozen rows (0.2 pts)")
                    total_score += 0.2
                else:
                    print(f"FAIL: Component 2 — Headers corrupted. Expected {expected_headers}, got {actual_headers}")
            else:
                print(f"FAIL: Component 2 — Freeze row < 3, headers not in frozen area")
        else:
            print(f"FAIL: Component 2 — No freeze panes set, cannot verify frozen header integrity")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Data preservation with freeze (0.2 points)
    # Verify the sheet still has data rows 3-502 and the freeze didn't corrupt anything.
    # Gated on freeze_panes being set.
    try:
        fp = ws.freeze_panes
        if fp is not None:
            max_row = ws.max_row
            if max_row >= 502:
                # Spot-check a data cell in row 3
                r3_val = ws.cell(row=3, column=1).value
                if r3_val is not None and isinstance(r3_val, str) and len(r3_val) > 0:
                    print(f"PASS: Component 3 — Data preserved ({max_row} rows, row 3 has '{r3_val}') (0.2 pts)")
                    total_score += 0.2
                else:
                    print(f"FAIL: Component 3 — Row 3 data missing or invalid: {r3_val}")
            else:
                print(f"FAIL: Component 3 — Expected >=502 rows, found {max_row}")
        else:
            print(f"FAIL: Component 3 — No freeze panes set, data integrity check skipped")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
