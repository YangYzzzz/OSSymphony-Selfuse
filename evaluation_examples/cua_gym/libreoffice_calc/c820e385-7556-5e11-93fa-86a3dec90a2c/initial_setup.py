"""
Initial Setup: Freeze top two rows in inventory spreadsheet
Task ID: calc_gg5_002
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_002'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Stock'

    # --- Row 1: Headers ---
    headers = ['Item', 'SKU', 'Quantity', 'Unit', 'Price', 'Reorder Level']
    header_font = Font(bold=True, size=11, name='Calibri')
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Row 2: Column unit descriptions ---
    units = ['(product name)', '(alphanumeric)', '(count)', '(measure)', '(USD)', '(count)']
    unit_font = Font(italic=True, size=10, color="808080", name='Calibri')
    unit_align = Alignment(horizontal="center", vertical="center")
    for col, u in enumerate(units, 1):
        cell = ws.cell(row=2, column=col, value=u)
        cell.font = unit_font
        cell.alignment = unit_align

    # --- Rows 3-502: 500 rows of inventory data ---
    categories = {
        'Electronics': [
            'Wireless Mouse', 'USB-C Hub', 'Bluetooth Speaker', 'Webcam HD',
            'Mechanical Keyboard', 'Monitor Stand', 'HDMI Cable', 'Laptop Sleeve',
            'Power Bank 20000mAh', 'Noise-Canceling Headphones', 'USB Flash Drive 64GB',
            'Ethernet Adapter', 'Wireless Charger', 'LED Desk Lamp', 'Surge Protector',
            'External SSD 1TB', 'Graphics Tablet', 'Microphone Stand', 'Cable Management Kit',
            'Screen Protector Pack',
        ],
        'Office Supplies': [
            'Ballpoint Pen Pack', 'Sticky Notes Assorted', 'Binder Clips Box',
            'File Folders 50ct', 'Whiteboard Markers', 'Desk Organizer', 'Paper Shredder',
            'Laminating Pouches', 'Rubber Bands Bag', 'Correction Tape',
            'Stapler Heavy Duty', 'Paper Clips Box', 'Highlighter Set', 'Desk Calendar 2026',
            'Label Maker Tape', 'Envelope Pack 100', 'Copy Paper Ream', 'Push Pins 200ct',
            'Scissors Stainless', 'Tape Dispenser',
        ],
        'Furniture': [
            'Ergonomic Chair', 'Standing Desk Converter', 'Bookshelf 5-Tier',
            'Filing Cabinet 3-Drawer', 'Monitor Arm Dual', 'Footrest Adjustable',
            'Desk Mat XL', 'Cable Tray Under-Desk', 'Whiteboard 48x36',
            'Cork Board Large', 'Storage Bins Set', 'Coat Rack Freestanding',
            'Side Table Rolling', 'Keyboard Tray', 'Document Holder',
            'Privacy Screen 24in', 'Wall Shelf Set', 'Drawer Organizer Tray',
            'Desk Riser Block', 'Cable Clips Adhesive',
        ],
        'Breakroom': [
            'Coffee Pods Variety', 'Paper Cups 100ct', 'Sugar Packets Box',
            'Creamer Singles', 'Napkins 500ct', 'Plastic Utensils Set',
            'Water Filter Pitcher', 'Microwave Cover', 'Paper Plates 200ct',
            'Hand Soap Refill', 'Paper Towels 6-Roll', 'Trash Bags 50gal',
            'Disinfectant Wipes', 'Air Freshener Spray', 'First Aid Kit',
            'Hand Sanitizer 32oz', 'Tea Bags Assorted', 'Stir Sticks 1000ct',
            'Snack Basket Organizer', 'Water Bottle Steel',
        ],
        'Safety': [
            'Safety Glasses Clear', 'Ear Plugs 200-Pair', 'Hard Hat White',
            'Hi-Vis Vest Orange', 'Nitrile Gloves Box', 'Face Shield Full',
            'Dust Mask N95 20ct', 'Fire Extinguisher 5lb', 'Caution Tape Roll',
            'Floor Sign Wet', 'Lockout Tagout Kit', 'Emergency Blanket',
            'Smoke Detector Battery', 'Exit Sign LED', 'Fire Alarm Pull',
            'Spill Kit Absorbent', 'Safety Harness', 'Reflective Tape Roll',
            'Eye Wash Station', 'First Responder Bag',
        ],
    }

    unit_options = ['each', 'box', 'pack', 'set', 'roll', 'case', 'bag', 'pair', 'ream', 'carton']

    all_items = []
    for cat, items in categories.items():
        for item in items:
            all_items.append((item, cat))

    sku_counter = 1000
    for r in range(3, 503):  # rows 3 through 502 (500 rows)
        idx = (r - 3) % len(all_items)
        item_name, cat = all_items[idx]
        # Add variant suffix for repeats
        variant = (r - 3) // len(all_items)
        if variant > 0:
            item_name = f"{item_name} v{variant + 1}"

        sku = f"SKU-{cat[:3].upper()}-{sku_counter}"
        sku_counter += 1
        quantity = random.randint(0, 950)
        unit = random.choice(unit_options)
        price = round(random.uniform(2.50, 499.99), 2)
        reorder = random.randint(5, 100)

        ws.cell(row=r, column=1, value=item_name)
        ws.cell(row=r, column=2, value=sku)
        ws.cell(row=r, column=3, value=quantity)
        ws.cell(row=r, column=4, value=unit)
        ws.cell(row=r, column=5, value=price)
        ws.cell(row=r, column=5).number_format = '$#,##0.00'
        ws.cell(row=r, column=6, value=reorder)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15

    # MUST NOT freeze panes - that is the task
    ws.freeze_panes = None

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
