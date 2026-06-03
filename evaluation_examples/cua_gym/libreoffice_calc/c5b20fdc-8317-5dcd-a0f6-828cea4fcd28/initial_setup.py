"""
Initial Setup: Grade book with Master roster for class sections task
Task ID: calc_edu_class_sections_sheets_007
Domain: libreoffice_calc
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_class_sections_sheets_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Master ---
    ws = wb.active
    ws.title = 'Master'

    # Headers
    headers = ['Student Name', 'Section', 'Score1', 'Score2', 'Score3']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 80 students data — 20 per section, interleaved for realism
    # (not grouped by section so the filter task is meaningful)
    students = [
        # Section A students
        ('Emma Sullivan',    'Section A', 88, 92, 85),
        ('Liam Nakamura',    'Section A', 76, 81, 79),
        ('Olivia Martinez',  'Section A', 95, 90, 93),
        ('Noah Patel',       'Section A', 82, 78, 84),
        ('Ava Thompson',     'Section A', 69, 74, 71),
        ('Ethan Williams',   'Section A', 91, 88, 94),
        ('Sophia Davis',     'Section A', 78, 83, 80),
        ('Mason Brown',      'Section A', 85, 87, 82),
        ('Isabella Garcia',  'Section A', 73, 69, 75),
        ('Lucas Wilson',     'Section A', 90, 93, 91),
        # Section B students
        ('Mia Johnson',      'Section B', 84, 79, 88),
        ('James Anderson',   'Section B', 71, 76, 73),
        ('Charlotte Lee',    'Section B', 97, 94, 96),
        ('Benjamin Taylor',  'Section B', 65, 70, 68),
        ('Amelia Harris',    'Section B', 88, 91, 86),
        ('Henry Jackson',    'Section B', 79, 82, 77),
        ('Harper White',     'Section B', 92, 89, 95),
        ('Sebastian Thomas', 'Section B', 74, 71, 76),
        ('Evelyn Robinson',  'Section B', 86, 88, 83),
        ('Jack Lewis',       'Section B', 68, 65, 70),
        # Section A continued
        ('Aiden Walker',     'Section A', 77, 80, 78),
        ('Scarlett Hall',    'Section A', 93, 96, 91),
        ('Jackson Allen',    'Section A', 81, 84, 79),
        ('Grace Young',      'Section A', 87, 85, 89),
        ('Carter King',      'Section A', 72, 68, 74),
        ('Chloe Scott',      'Section A', 94, 91, 96),
        ('Jayden Green',     'Section A', 80, 76, 82),
        ('Zoey Adams',       'Section A', 89, 92, 87),
        ('Wyatt Baker',      'Section A', 75, 78, 73),
        ('Lily Nelson',      'Section A', 96, 93, 98),
        # Section C students
        ('Dylan Carter',     'Section C', 83, 86, 81),
        ('Nora Mitchell',    'Section C', 70, 67, 72),
        ('Elijah Perez',     'Section C', 91, 94, 89),
        ('Luna Roberts',     'Section C', 78, 75, 80),
        ('Grayson Turner',   'Section C', 86, 89, 84),
        ('Penelope Phillips','Section C', 62, 67, 65),
        ('Isaiah Campbell',  'Section C', 93, 90, 95),
        ('Riley Parker',     'Section C', 77, 80, 75),
        ('Eli Evans',        'Section C', 89, 85, 91),
        ('Layla Edwards',    'Section C', 74, 71, 76),
        # Section B continued
        ('Julian Collins',   'Section B', 82, 85, 80),
        ('Zoey Stewart',     'Section B', 90, 87, 92),
        ('Mateo Sanchez',    'Section B', 67, 72, 69),
        ('Hazel Morris',     'Section B', 95, 98, 93),
        ('Leo Rogers',       'Section B', 76, 73, 78),
        ('Violet Reed',      'Section B', 88, 91, 86),
        ('Gabriel Cook',     'Section B', 73, 70, 75),
        ('Aurora Morgan',    'Section B', 94, 97, 92),
        ('Levi Bell',        'Section B', 69, 66, 71),
        ('Stella Murphy',    'Section B', 85, 88, 83),
        # Section C continued
        ('Owen Bailey',      'Section C', 88, 84, 90),
        ('Hannah Rivera',    'Section C', 71, 75, 73),
        ('Samuel Cooper',    'Section C', 96, 92, 98),
        ('Addison Richardson','Section C', 80, 83, 78),
        ('Joseph Cox',       'Section C', 85, 88, 83),
        ('Nora Howard',      'Section C', 66, 63, 68),
        ('David Ward',       'Section C', 92, 89, 94),
        ('Leah Torres',      'Section C', 79, 76, 81),
        ('John Peterson',    'Section C', 87, 90, 85),
        ('Natalie Gray',     'Section C', 73, 70, 75),
        # Section D students
        ('Andrew James',     'Section D', 84, 87, 82),
        ('Audrey Watson',    'Section D', 91, 88, 93),
        ('Ryan Brooks',      'Section D', 77, 74, 79),
        ('Brooklyn Kelly',   'Section D', 95, 92, 97),
        ('Christian Sanders','Section D', 70, 73, 68),
        ('Aria Price',       'Section D', 88, 91, 86),
        ('Jonathan Bennett', 'Section D', 75, 72, 77),
        ('Savannah Wood',    'Section D', 93, 96, 91),
        ('Nathan Barnes',    'Section D', 68, 65, 70),
        ('Paisley Ross',     'Section D', 86, 89, 84),
        ('Aaron Henderson',  'Section D', 79, 82, 77),
        ('Alexa Coleman',    'Section D', 97, 94, 99),
        ('Tyler Jenkins',    'Section D', 72, 69, 74),
        ('Stella Perry',     'Section D', 90, 93, 88),
        ('Nicholas Powell',  'Section D', 83, 80, 85),
        ('Bella Patterson',  'Section D', 76, 79, 74),
        ('Adam Hughes',      'Section D', 94, 91, 96),
        ('Elena Flores',     'Section D', 69, 72, 67),
        ('Caleb Washington', 'Section D', 87, 84, 89),
        ('Serenity Butler',  'Section D', 81, 78, 83),
    ]

    for r, row_data in enumerate(students, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Master with {len(students)} students')


create_initial()
