"""
Reward Script: Organize grade book by creating separate section sheets
Task ID: calc_edu_class_sections_sheets_007
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Four section sheets created (Section A, B, C, D)
  Component 2 (0.20): Each section sheet has correct header row (copied from Master row 1)
  Component 3 (0.30): Each section sheet contains exactly the correct students for that section
  Component 4 (0.20): Student order in section sheets matches relative order in Master
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_class_sections_sheets_007'

EXPECTED_SECTIONS = ['Section A', 'Section B', 'Section C', 'Section D']
EXPECTED_HEADERS = ('Student Name', 'Section', 'Score1', 'Score2', 'Score3')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify Master sheet exists (precondition gate)
    if 'Master' not in wb.sheetnames:
        print("CRITICAL: 'Master' sheet missing — file may be corrupted")
        print("REWARD: 0.0")
        return 0.0

    # Read Master sheet to get expected data
    ws_master = wb['Master']
    master_header = tuple(ws_master.cell(row=1, column=c).value for c in range(1, 6))

    # Build expected mapping: section_name -> list of rows (as tuples) in master order
    expected_section_rows = {s: [] for s in EXPECTED_SECTIONS}
    for row_idx in range(2, ws_master.max_row + 1):
        row_data = tuple(ws_master.cell(row=row_idx, column=c).value for c in range(1, 6))
        section_val = row_data[1]  # Column B
        if section_val in expected_section_rows:
            expected_section_rows[section_val].append(row_data)

    # Component 1: Four section sheets exist — Section A, B, C, D (0.30 points)
    # This FAILS on initial (only 'Master' exists) and PASSES on golden
    try:
        sections_present = [s for s in EXPECTED_SECTIONS if s in wb.sheetnames]
        if len(sections_present) == 4:
            print(f"PASS: Component 1 — All 4 section sheets present: {sections_present} (0.30 pts)")
            total_score += 0.30
        else:
            missing = [s for s in EXPECTED_SECTIONS if s not in wb.sheetnames]
            print(f"FAIL: Component 1 — Missing section sheets: {missing}. Present: {sections_present}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Each section sheet has correct header row (copied from Master row 1) (0.20 points)
    # This FAILS on initial (sheets don't exist) and PASSES on golden
    try:
        header_pass_count = 0
        header_issues = []
        for section in EXPECTED_SECTIONS:
            if section not in wb.sheetnames:
                header_issues.append(f"'{section}' sheet missing")
                continue
            ws_sec = wb[section]
            actual_header = tuple(ws_sec.cell(row=1, column=c).value for c in range(1, 6))
            if actual_header == master_header:
                header_pass_count += 1
            else:
                header_issues.append(f"'{section}' header {actual_header!r} != master {master_header!r}")

        if header_pass_count == 4:
            print(f"PASS: Component 2 — All 4 section sheets have correct header row (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 — {4 - header_pass_count}/4 headers correct. Issues: {header_issues}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Each section sheet contains exactly the correct students for that section (0.30 points)
    # "Correct" = only students belonging to that section, with correct row count
    # This FAILS on initial (sheets don't exist) and PASSES on golden
    try:
        data_pass_count = 0
        data_issues = []
        for section in EXPECTED_SECTIONS:
            if section not in wb.sheetnames:
                data_issues.append(f"'{section}' sheet missing")
                continue
            ws_sec = wb[section]
            # Data rows start at row 2 (row 1 is header)
            actual_rows = []
            for row_idx in range(2, ws_sec.max_row + 1):
                row_data = tuple(ws_sec.cell(row=row_idx, column=c).value for c in range(1, 6))
                # Only count non-empty rows
                if any(v is not None for v in row_data):
                    actual_rows.append(row_data)

            expected_rows = expected_section_rows[section]
            expected_count = len(expected_rows)
            actual_count = len(actual_rows)

            if actual_count != expected_count:
                data_issues.append(f"'{section}' has {actual_count} data rows, expected {expected_count}")
                continue

            # Verify all rows have correct section value in column B
            wrong_section_rows = [r for r in actual_rows if r[1] != section]
            if wrong_section_rows:
                data_issues.append(f"'{section}' has {len(wrong_section_rows)} rows with wrong section value")
                continue

            # Verify exact data content matches expected
            if set(actual_rows) == set(expected_rows):
                data_pass_count += 1
            else:
                data_issues.append(f"'{section}' data content mismatch")

        if data_pass_count == 4:
            print(f"PASS: Component 3 — All 4 section sheets have correct student data (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 3 — {data_pass_count}/4 sections have correct data. Issues: {data_issues}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Student order matches relative order in Master (0.20 points)
    # This FAILS on initial (sheets don't exist) and PASSES on golden
    try:
        order_pass_count = 0
        order_issues = []
        for section in EXPECTED_SECTIONS:
            if section not in wb.sheetnames:
                order_issues.append(f"'{section}' sheet missing")
                continue
            ws_sec = wb[section]
            # Get student names (column A) from section sheet rows 2+
            actual_names = []
            for row_idx in range(2, ws_sec.max_row + 1):
                name = ws_sec.cell(row=row_idx, column=1).value
                if name is not None:
                    actual_names.append(name)

            # Get expected order from master
            expected_names = [r[0] for r in expected_section_rows[section]]

            if actual_names == expected_names:
                order_pass_count += 1
            else:
                order_issues.append(f"'{section}' order mismatch: actual={actual_names[:3]}... expected={expected_names[:3]}...")

        if order_pass_count == 4:
            print(f"PASS: Component 4 — All 4 section sheets have correct student order (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 4 — {order_pass_count}/4 sections have correct order. Issues: {order_issues}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
