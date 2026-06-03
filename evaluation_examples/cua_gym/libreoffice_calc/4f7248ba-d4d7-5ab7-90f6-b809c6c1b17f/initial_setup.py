"""
Initial Setup: Vendor AP Aging Report — pre-task state (no aging bucket columns)
Task ID: calc_fin_vendor_aging_009
Domain: libreoffice_calc

Creates an AP sheet with 49 vendor invoices (rows 2-50).
Columns A-E only; no aging bucket columns (F-I), no row-51 totals, no charts.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date, timedelta
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_vendor_aging_009'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'AP'

    # --- Headers (Row 1) ---
    headers = ['Vendor', 'Invoice#', 'Invoice Date', 'Amount', 'Days Outstanding']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # --- Realistic vendor invoice data (49 rows, varied days outstanding) ---
    vendors = [
        'Apex Office Supplies', 'Brightway Logistics', 'Cascade Tech Solutions',
        'Delta Printing Co', 'Eastgate Marketing', 'Frontier Hardware',
        'Global Freight Services', 'Harbor Consulting', 'Inland Distribution',
        'Jupiter Software Inc', 'Kinetic Staffing', 'Lumina Design Studio',
        'Metro Cleaning Services', 'Nexus Data Corp', 'Omega Legal Partners',
        'Pacific Rim Trading', 'Quantum Analytics', 'Redwood Catering',
        'Stellar Security', 'Titan Engineering',
    ]

    invoices = [
        # (vendor, invoice_num, invoice_date, amount, days_outstanding)
        ('Apex Office Supplies',    'INV-2025-0041', date(2025, 2, 10),  1250.00,  21),
        ('Brightway Logistics',     'INV-2025-0087', date(2025, 1, 25),  8740.50,  37),
        ('Cascade Tech Solutions',  'INV-2025-0023', date(2025, 1, 8),   4300.00,  54),
        ('Delta Printing Co',       'INV-2025-0104', date(2024, 12, 22), 980.75,   70),
        ('Eastgate Marketing',      'INV-2025-0116', date(2024, 12, 5),  6150.00,  87),
        ('Frontier Hardware',       'INV-2025-0055', date(2024, 11, 10), 3420.25,  112),
        ('Global Freight Services', 'INV-2025-0132', date(2025, 2, 18),  2100.00,  13),
        ('Harbor Consulting',       'INV-2025-0078', date(2025, 2, 1),   9500.00,  30),
        ('Inland Distribution',     'INV-2025-0091', date(2025, 1, 14),  1875.50,  47),
        ('Jupiter Software Inc',    'INV-2025-0029', date(2024, 12, 30), 5600.00,  62),
        ('Kinetic Staffing',        'INV-2025-0145', date(2024, 12, 12), 11200.00, 79),
        ('Lumina Design Studio',    'INV-2025-0063', date(2024, 11, 25), 2350.00,  96),
        ('Metro Cleaning Services', 'INV-2025-0017', date(2025, 2, 22),  450.00,   9),
        ('Nexus Data Corp',         'INV-2025-0098', date(2025, 2, 5),   7800.00,  26),
        ('Omega Legal Partners',    'INV-2025-0082', date(2025, 1, 20),  3150.00,  41),
        ('Pacific Rim Trading',     'INV-2025-0111', date(2025, 1, 3),   6900.50,  58),
        ('Quantum Analytics',       'INV-2025-0044', date(2024, 12, 18), 4250.00,  74),
        ('Redwood Catering',        'INV-2025-0137', date(2024, 12, 1),  890.25,   91),
        ('Stellar Security',        'INV-2025-0006', date(2024, 11, 14), 5375.00,  107),
        ('Titan Engineering',       'INV-2025-0120', date(2025, 2, 14),  12400.00, 17),
        ('Apex Office Supplies',    'INV-2025-0160', date(2025, 2, 2),   675.00,   29),
        ('Brightway Logistics',     'INV-2025-0172', date(2025, 1, 17),  3100.00,  44),
        ('Cascade Tech Solutions',  'INV-2025-0058', date(2024, 12, 28), 8200.00,  64),
        ('Delta Printing Co',       'INV-2025-0193', date(2024, 12, 10), 1540.00,  82),
        ('Eastgate Marketing',      'INV-2025-0210', date(2024, 11, 22), 7350.00,  99),
        ('Frontier Hardware',       'INV-2025-0185', date(2025, 2, 25),  2200.00,  6),
        ('Global Freight Services', 'INV-2025-0047', date(2025, 2, 10),  4800.00,  21),
        ('Harbor Consulting',       'INV-2025-0228', date(2025, 1, 26),  6050.00,  36),
        ('Inland Distribution',     'INV-2025-0203', date(2025, 1, 9),   3380.50,  52),
        ('Jupiter Software Inc',    'INV-2025-0166', date(2024, 12, 23), 9100.00,  69),
        ('Kinetic Staffing',        'INV-2025-0241', date(2024, 12, 7),  5200.00,  85),
        ('Lumina Design Studio',    'INV-2025-0154', date(2024, 11, 18), 1700.00,  103),
        ('Metro Cleaning Services', 'INV-2025-0219', date(2025, 2, 20),  510.00,   11),
        ('Nexus Data Corp',         'INV-2025-0177', date(2025, 2, 6),   8900.00,  25),
        ('Omega Legal Partners',    'INV-2025-0235', date(2025, 1, 22),  4400.00,  39),
        ('Pacific Rim Trading',     'INV-2025-0199', date(2025, 1, 5),   7250.00,  56),
        ('Quantum Analytics',       'INV-2025-0142', date(2024, 12, 20), 3600.00,  72),
        ('Redwood Catering',        'INV-2025-0256', date(2024, 12, 3),  1020.50,  89),
        ('Stellar Security',        'INV-2025-0125', date(2024, 11, 15), 6480.00,  106),
        ('Titan Engineering',       'INV-2025-0268', date(2025, 2, 16),  10750.00, 15),
        ('Apex Office Supplies',    'INV-2025-0280', date(2025, 2, 3),   825.00,   28),
        ('Brightway Logistics',     'INV-2025-0274', date(2025, 1, 18),  5150.00,  43),
        ('Cascade Tech Solutions',  'INV-2025-0289', date(2024, 12, 29), 7650.00,  63),
        ('Delta Printing Co',       'INV-2025-0292', date(2024, 12, 11), 2100.00,  81),
        ('Eastgate Marketing',      'INV-2025-0301', date(2024, 11, 23), 8900.00,  98),
        ('Frontier Hardware',       'INV-2025-0315', date(2025, 2, 27),  1350.00,  4),
        ('Global Freight Services', 'INV-2025-0322', date(2025, 2, 12),  3700.00,  19),
        ('Harbor Consulting',       'INV-2025-0308', date(2025, 1, 28),  6800.00,  33),
        ('Inland Distribution',     'INV-2025-0296', date(2025, 1, 11),  4450.00,  50),
    ]

    for r, (vendor, inv_num, inv_date, amount, days) in enumerate(invoices, 2):
        ws.cell(row=r, column=1, value=vendor)
        ws.cell(row=r, column=2, value=inv_num)
        date_cell = ws.cell(row=r, column=3, value=inv_date)
        date_cell.number_format = 'yyyy-mm-dd'
        amount_cell = ws.cell(row=r, column=4, value=amount)
        amount_cell.number_format = '$#,##0.00'
        ws.cell(row=r, column=5, value=days)

    # --- Column widths for readability ---
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: AP')
    print(f'  Rows: 1 header + 49 data rows (rows 2-50)')
    print(f'  Columns: A-E only (no aging bucket columns)')
    print(f'  Days distribution: 0-30: ~14, 31-60: ~13, 61-90: ~11, 90+: ~11')

create_initial()
