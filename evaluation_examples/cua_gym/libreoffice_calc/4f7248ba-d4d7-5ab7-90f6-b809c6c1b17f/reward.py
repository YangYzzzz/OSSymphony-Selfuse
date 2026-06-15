"""
Reward Script: AP Aging Report with Bucket Columns and Pie Chart
Task ID: calc_fin_vendor_aging_009
Domain: libreoffice_calc
Scoring:
  Component 1: Aging bucket headers in F1:I1 (0.15 pts)
  Component 2: IF formulas in F2:I50 for aging buckets (0.35 pts)
  Component 3: SUM formulas in F51:I51 (0.15 pts)
  Component 4: Row 51 bold and currency formatted (0.15 pts)
  Component 5: Pie chart present with title 'AP Aging Distribution' (0.20 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fin_vendor_aging_009'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the AP sheet exists
    if 'AP' not in wb.sheetnames:
        print("CRITICAL: Sheet 'AP' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['AP']

    # Component 1: Aging bucket column headers in F1:I1 (0.15 points)
    # Expected: F1='0-30 Days', G1='31-60 Days', H1='61-90 Days', I1='90+ Days'
    try:
        expected_headers = {
            'F': '0-30 Days',
            'G': '31-60 Days',
            'H': '61-90 Days',
            'I': '90+ Days',
        }
        headers_correct = 0
        headers_found = {}
        for col, expected in expected_headers.items():
            actual = ws[f'{col}1'].value
            headers_found[col] = actual
            if actual and str(actual).strip() == expected:
                headers_correct += 1

        if headers_correct == 4:
            print(f"PASS: Component 1 — All 4 aging bucket headers correct in F1:I1 (0.15 pts)")
            total_score += 0.15
        elif headers_correct >= 2:
            print(f"PARTIAL: Component 1 — {headers_correct}/4 aging bucket headers correct, found: {headers_found}")
        else:
            print(f"FAIL: Component 1 — Expected aging bucket headers in F1:I1, found: {headers_found}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: IF formulas in F2:I50 for aging buckets (0.35 points)
    # F: =IF(E{n}<=30,D{n},0)
    # G: =IF(AND(E{n}>30,E{n}<=60),D{n},0)
    # H: =IF(AND(E{n}>61,E{n}<=90),D{n},0) -- note: >61 per task context
    # I: =IF(E{n}>90,D{n},0)
    try:
        formula_errors = []
        f_correct = 0
        g_correct = 0
        h_correct = 0
        i_correct = 0

        for row in range(2, 51):
            f_val = ws[f'F{row}'].value
            g_val = ws[f'G{row}'].value
            h_val = ws[f'H{row}'].value
            i_val = ws[f'I{row}'].value

            # Check F column: =IF(E{row}<=30,D{row},0)
            if f_val and isinstance(f_val, str):
                normalized = f_val.upper().replace(' ', '')
                expected_f = f'=IF(E{row}<=30,D{row},0)'.upper()
                if normalized == expected_f:
                    f_correct += 1
                elif f'E{row}' in normalized and '<=30' in normalized:
                    f_correct += 1  # accept minor variation
            else:
                formula_errors.append(f'F{row}: expected formula, got {repr(f_val)}')

            # Check G column: =IF(AND(E{row}>30,E{row}<=60),D{row},0)
            if g_val and isinstance(g_val, str):
                normalized = g_val.upper().replace(' ', '')
                if f'E{row}' in normalized and '>30' in normalized and '<=60' in normalized:
                    g_correct += 1
            else:
                formula_errors.append(f'G{row}: expected formula, got {repr(g_val)}')

            # Check H column: =IF(AND(E{row}>61,E{row}<=90),D{row},0) or >60 variant
            if h_val and isinstance(h_val, str):
                normalized = h_val.upper().replace(' ', '')
                # Accept either >60 or >61 (task context says >61 but >60 is also reasonable)
                if f'E{row}' in normalized and ('<=90' in normalized or '<91' in normalized):
                    h_correct += 1
            else:
                formula_errors.append(f'H{row}: expected formula, got {repr(h_val)}')

            # Check I column: =IF(E{row}>90,D{row},0)
            if i_val and isinstance(i_val, str):
                normalized = i_val.upper().replace(' ', '')
                if f'E{row}' in normalized and '>90' in normalized:
                    i_correct += 1
            else:
                formula_errors.append(f'I{row}: expected formula, got {repr(i_val)}')

        total_formula_rows = 49  # rows 2-50
        f_pct = f_correct / total_formula_rows
        g_pct = g_correct / total_formula_rows
        h_pct = h_correct / total_formula_rows
        i_pct = i_correct / total_formula_rows
        avg_pct = (f_pct + g_pct + h_pct + i_pct) / 4

        # Award points proportionally based on formula coverage
        component2_score = round(avg_pct * 0.35, 4)
        if avg_pct >= 0.99:
            print(f"PASS: Component 2 — IF formulas in F2:I50 all correct ({f_correct}/{total_formula_rows} F, {g_correct}/{total_formula_rows} G, {h_correct}/{total_formula_rows} H, {i_correct}/{total_formula_rows} I) (0.35 pts)")
        elif avg_pct >= 0.5:
            print(f"PARTIAL: Component 2 — IF formulas partially correct: F={f_correct}, G={g_correct}, H={h_correct}, I={i_correct} out of {total_formula_rows}")
        else:
            print(f"FAIL: Component 2 — IF formulas mostly missing or wrong: F={f_correct}, G={g_correct}, H={h_correct}, I={i_correct}")
            if formula_errors:
                print(f"  First errors: {formula_errors[:3]}")
        total_score += component2_score

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: SUM formulas in F51:I51 (0.15 points)
    # F51=SUM(F2:F50), G51=SUM(G2:G50), H51=SUM(H2:H50), I51=SUM(I2:I50)
    try:
        sum_cols_correct = 0
        sum_found = {}
        for col, expected_range in [('F', 'F2:F50'), ('G', 'G2:G50'), ('H', 'H2:H50'), ('I', 'I2:I50')]:
            cell_val = ws[f'{col}51'].value
            sum_found[col] = cell_val
            if cell_val and isinstance(cell_val, str):
                normalized = cell_val.upper().replace(' ', '')
                expected = f'=SUM({expected_range})'.upper()
                if normalized == expected:
                    sum_cols_correct += 1

        if sum_cols_correct == 4:
            print(f"PASS: Component 3 — All 4 SUM formulas in row 51 correct (0.15 pts)")
            total_score += 0.15
        elif sum_cols_correct >= 2:
            partial = round(sum_cols_correct / 4 * 0.15, 4)
            print(f"PARTIAL: Component 3 — {sum_cols_correct}/4 SUM formulas correct, found: {sum_found}")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Expected SUM formulas in F51:I51, found: {sum_found}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Row 51 cells F51:I51 bold and currency formatted (0.15 points)
    # Bold font and number_format='$#,##0.00' (or similar currency format)
    try:
        bold_count = 0
        currency_count = 0
        style_details = {}
        for col in ['F', 'G', 'H', 'I']:
            cell = ws[f'{col}51']
            is_bold = cell.font.bold is True
            fmt = cell.number_format
            # Accept various currency formats: $#,##0.00, $#,##0, etc.
            is_currency = (
                fmt is not None and (
                    '$' in str(fmt) or
                    '0.00' in str(fmt) or
                    '#,##0' in str(fmt)
                )
            )
            style_details[col] = {'bold': is_bold, 'format': fmt}
            if is_bold:
                bold_count += 1
            if is_currency:
                currency_count += 1

        all_bold = (bold_count == 4)
        all_currency = (currency_count == 4)

        if all_bold and all_currency:
            print(f"PASS: Component 4 — Row 51 (F-I) all bold and currency formatted (0.15 pts)")
            total_score += 0.15
        elif all_bold:
            print(f"PARTIAL: Component 4 — Row 51 bold={bold_count}/4, currency={currency_count}/4. Details: {style_details}")
            total_score += 0.07
        elif all_currency:
            print(f"PARTIAL: Component 4 — Row 51 bold={bold_count}/4, currency={currency_count}/4. Details: {style_details}")
            total_score += 0.07
        else:
            print(f"FAIL: Component 4 — Row 51 bold={bold_count}/4, currency={currency_count}/4. Details: {style_details}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Pie chart present with title 'AP Aging Distribution' (0.20 points)
    # The chart should be a PieChart using F51:I51 data
    try:
        charts = ws._charts
        if not charts:
            print("FAIL: Component 5 — No chart found in the AP sheet")
        else:
            # Find a pie chart
            pie_chart = None
            for c in charts:
                if type(c).__name__ == 'PieChart':
                    pie_chart = c
                    break

            if pie_chart is None:
                print(f"FAIL: Component 5 — Found {len(charts)} chart(s) but none is a PieChart")
            else:
                # Check chart title
                chart_title = None
                try:
                    chart_title = pie_chart.title.tx.rich.p[0].r[0].t
                except Exception:
                    try:
                        chart_title = str(pie_chart.title)
                    except Exception:
                        chart_title = None

                title_ok = (
                    chart_title is not None and
                    'AP Aging Distribution' in str(chart_title)
                )

                # Check that chart uses F51:I51 data (series reference)
                data_ok = False
                series_refs = []
                try:
                    for ser in pie_chart.series:
                        if ser.val and ser.val.numRef:
                            series_refs.append(ser.val.numRef.f)
                    # Check if series references include row 51 cells from columns F, G, H, or I
                    # Refs can be like "'AP'!$F$51" or "AP!F51" — strip all special chars
                    row51_buckets = {'F51', 'G51', 'H51', 'I51'}
                    found_bucket_refs = set()
                    for ref in series_refs:
                        # Remove quotes, dollar signs, spaces for comparison
                        ref_clean = ref.replace("'", '').replace('$', '').replace(' ', '').upper()
                        for bucket in row51_buckets:
                            if bucket in ref_clean:
                                found_bucket_refs.add(bucket)
                    # Also accept: single data reference spanning F51:I51
                    if not found_bucket_refs:
                        for ref in series_refs:
                            ref_clean = ref.replace("'", '').replace('$', '').replace(' ', '').upper()
                            if 'F51' in ref_clean and 'I51' in ref_clean:
                                found_bucket_refs = row51_buckets
                                break
                    data_ok = len(found_bucket_refs) >= 3  # at least 3 of 4 bucket columns
                except Exception as ex:
                    print(f"  Chart data check error: {ex}")

                if title_ok and data_ok:
                    print(f"PASS: Component 5 — PieChart found with title '{chart_title}' using AP totals data (0.20 pts)")
                    total_score += 0.20
                elif title_ok:
                    print(f"PARTIAL: Component 5 — PieChart title OK ('{chart_title}') but data refs don't clearly reference row 51: {series_refs}")
                    total_score += 0.10
                elif data_ok:
                    print(f"PARTIAL: Component 5 — PieChart data refs OK but title wrong/missing (found: {repr(chart_title)})")
                    total_score += 0.10
                else:
                    print(f"FAIL: Component 5 — PieChart title='{chart_title}', data OK={data_ok}, refs={series_refs if 'series_refs' in dir() else 'unknown'}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
