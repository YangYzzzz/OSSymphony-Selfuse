"""
Initial Setup: Merge cells B3:D3, center-align, double bottom border, light orange background
Task ID: calc_gg3_036
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_036'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Invoice Sheet ---
    ws = wb.active
    ws.title = 'Invoice'

    # Company header area
    ws['B1'] = 'Northwind Trading Co.'
    ws['B1'].font = Font(name='Arial', size=14, bold=True)
    ws['B2'] = '742 Evergreen Terrace, Suite 200, Portland, OR 97201'
    ws['B2'].font = Font(name='Arial', size=9, color='666666')

    # Invoice number row - B3 has text, C3 and D3 are empty (NO merge, NO formatting)
    ws['B3'] = 'Invoice Number: INV-2024-0056'
    ws['B3'].font = Font(name='Arial', size=11)
    # C3 and D3 intentionally left empty

    # Invoice metadata
    ws['B4'] = 'Date:'
    ws['C4'] = '2024-11-15'
    ws['B5'] = 'Due Date:'
    ws['C5'] = '2024-12-15'
    ws['B6'] = 'Payment Terms:'
    ws['C6'] = 'Net 30'

    # Bill To section
    ws['B8'] = 'Bill To:'
    ws['B8'].font = Font(name='Arial', size=10, bold=True)
    ws['B9'] = 'Pinnacle Solutions Inc.'
    ws['B10'] = 'Attn: Rebecca Torres'
    ws['B11'] = '1200 Market Street, Floor 8'
    ws['B12'] = 'San Francisco, CA 94103'

    # Line items header
    headers = ['Item', 'Description', 'Qty', 'Unit Price', 'Amount']
    header_cols = [2, 3, 4, 5, 6]  # B through F
    thin_border = Side(style='thin', color='000000')
    for col, header in zip(header_cols, headers):
        cell = ws.cell(row=14, column=col, value=header)
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.border = Border(bottom=thin_border)
        cell.alignment = Alignment(horizontal='center')

    # Line items data
    items = [
        ['WD-1001', 'Web Development - Homepage Redesign', 40, 125.00, 5000.00],
        ['WD-1002', 'Web Development - Product Pages (x5)', 60, 125.00, 7500.00],
        ['GD-2001', 'Graphic Design - Logo Package', 1, 2500.00, 2500.00],
        ['GD-2002', 'Graphic Design - Marketing Banners (x8)', 8, 175.00, 1400.00],
        ['SEO-3001', 'SEO Audit & Optimization', 20, 95.00, 1900.00],
        ['CW-4001', 'Copywriting - Website Content', 15, 85.00, 1275.00],
        ['PM-5001', 'Project Management', 25, 110.00, 2750.00],
        ['QA-6001', 'Quality Assurance & Testing', 12, 90.00, 1080.00],
        ['TR-7001', 'Training & Handoff Sessions', 4, 200.00, 800.00],
        ['SP-8001', 'Hosting Setup & Configuration', 1, 500.00, 500.00],
    ]

    for r, row_data in enumerate(items, 15):
        for c, val in zip(header_cols, row_data):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name='Arial', size=10)
            if isinstance(val, float):
                cell.number_format = '#,##0.00'

    # Subtotal, Tax, Total
    ws.cell(row=26, column=5, value='Subtotal:').font = Font(name='Arial', size=10, bold=True)
    ws.cell(row=26, column=6, value=24705.00).number_format = '$#,##0.00'
    ws.cell(row=27, column=5, value='Tax (8.5%):').font = Font(name='Arial', size=10)
    ws.cell(row=27, column=6, value=2099.93).number_format = '$#,##0.00'
    ws.cell(row=28, column=5, value='Total Due:').font = Font(name='Arial', size=11, bold=True)
    ws.cell(row=28, column=6, value=26804.93).number_format = '$#,##0.00'

    # Notes
    ws['B30'] = 'Notes:'
    ws['B30'].font = Font(name='Arial', size=10, bold=True)
    ws['B31'] = 'Please make payment via bank transfer to the account details provided separately.'
    ws['B32'] = 'Thank you for your business!'

    # Column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 38
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
