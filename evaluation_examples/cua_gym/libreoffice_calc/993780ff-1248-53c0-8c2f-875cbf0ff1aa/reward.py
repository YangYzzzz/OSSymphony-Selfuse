"""
Reward Script: Merge B3:D3 on Invoice sheet, center-align, double bottom border, light orange fill
Task ID: calc_gg3_036
Domain: libreoffice_calc
Scoring:
  Component 1: B3:D3 merged AND text preserved (0.30)
  Component 2: Horizontal center alignment (0.25)
  Component 3: Double-line bottom border (0.25)
  Component 4: Light orange background #FFE4B5 (0.20)
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_036'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice documents."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Invoice' sheet must exist
    if 'Invoice' not in wb.sheetnames:
        print("CRITICAL: 'Invoice' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Invoice']

    # Component 1: Cells B3:D3 are merged AND text is preserved (0.30 points)
    # The merge is task-introduced; text preservation is verified as a sub-condition
    try:
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        b3_val = ws['B3'].value
        expected_text = 'Invoice Number: INV-2024-0056'
        is_merged = 'B3:D3' in merged_ranges
        text_ok = b3_val and str(b3_val).strip() == expected_text
        if is_merged and text_ok:
            print(f"PASS: Component 1 - B3:D3 merged and text preserved: '{b3_val}' (0.30 pts)")
            total_score += 0.30
        elif is_merged and not text_ok:
            print(f"PARTIAL: Component 1 - B3:D3 merged but text lost. Found: {repr(b3_val)} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 - B3:D3 not merged. Merged ranges: {merged_ranges}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Horizontal center alignment on B3 (0.25 points)
    try:
        h_align = ws['B3'].alignment.horizontal
        if h_align == 'center':
            print(f"PASS: Component 2 - Horizontal alignment is 'center' (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 - Expected horizontal='center', found: {repr(h_align)}")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Double-line bottom border on B3 (0.25 points)
    try:
        bottom_style = ws['B3'].border.bottom.style
        if bottom_style == 'double':
            print(f"PASS: Component 3 - Bottom border style is 'double' (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 3 - Expected bottom border 'double', found: {repr(bottom_style)}")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Background fill is light orange #FFE4B5 (0.20 points)
    try:
        fill_type = ws['B3'].fill.fill_type
        fg_rgb = ws['B3'].fill.fgColor.rgb if ws['B3'].fill.fgColor else None
        # Expected ARGB: FFFFE4B5 (FF alpha + FFE4B5 color)
        if fill_type == 'solid' and fg_rgb and fg_rgb.upper() == 'FFFFE4B5':
            print(f"PASS: Component 4 - Background is solid #FFE4B5 (ARGB: {fg_rgb}) (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 4 - Expected solid fill FFFFE4B5, found fill_type={repr(fill_type)}, fgColor={repr(fg_rgb)}")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
