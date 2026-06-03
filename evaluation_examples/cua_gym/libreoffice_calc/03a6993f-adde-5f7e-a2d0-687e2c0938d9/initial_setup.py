"""
Initial Setup: Filtered dataset with SUM formula that needs fixing to SUBTOTAL
Task ID: calc_tbl_024
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.filters import AutoFilter, FilterColumn, Filters

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_024'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    # --- Headers ---
    headers = ["Region", "Salesperson", "Product", "Amount"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # --- Data rows (rows 2-100, 99 data rows) ---
    regions = ["East", "West", "North", "South"]
    salesperson_pool = {
        "East": ["Sarah Chen", "Marcus Johnson", "Priya Patel", "David Kim", "Emily Rodriguez"],
        "West": ["James Wilson", "Lisa Nakamura", "Carlos Reyes", "Amanda Foster", "Ryan O'Brien"],
        "North": ["Michelle Lee", "Thomas Anderson", "Fatima Hassan", "Kevin Murphy", "Jessica Wang"],
        "South": ["Robert Taylor", "Olivia Brown", "Daniel Garcia", "Hannah Stewart", "Christopher Ng"],
    }
    products = [
        "Laptop Pro 15", "Wireless Mouse", "USB-C Hub", "Monitor 27\"",
        "Mechanical Keyboard", "Webcam HD", "Docking Station", "SSD 1TB",
        "RAM 32GB Kit", "Headset Elite", "Tablet 10\"", "Printer LaserJet",
    ]

    data_font = Font(name="Calibri", size=11)
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    amount_format = '$#,##0.00'

    for r in range(2, 101):
        region = random.choice(regions)
        salesperson = random.choice(salesperson_pool[region])
        product = random.choice(products)
        amount = round(random.uniform(150.00, 9500.00), 2)

        ws.cell(row=r, column=1, value=region).font = data_font
        ws.cell(row=r, column=2, value=salesperson).font = data_font
        ws.cell(row=r, column=3, value=product).font = data_font
        amt_cell = ws.cell(row=r, column=4, value=amount)
        amt_cell.font = data_font
        amt_cell.number_format = amount_format

        for c in range(1, 5):
            ws.cell(row=r, column=c).border = data_border

    # --- Total row (row 101) ---
    label_cell = ws.cell(row=101, column=3, value="Total:")
    label_cell.font = Font(name="Calibri", size=11, bold=True)
    label_cell.alignment = Alignment(horizontal="right")
    label_cell.border = data_border

    total_cell = ws.cell(row=101, column=4, value="=SUM(D2:D100)")
    total_cell.font = Font(name="Calibri", size=11, bold=True)
    total_cell.number_format = amount_format
    total_cell.border = data_border

    # --- Column widths ---
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 14

    # --- Auto-filter (definition only; actual filtering happens in LibreOffice) ---
    ws.auto_filter.ref = "A1:D100"
    # Add filter criteria for Region = East
    col_filter = FilterColumn(colId=0)
    col_filter.filters = Filters(filter=["East"])
    ws.auto_filter.filterColumn.append(col_filter)

    # --- Freeze header row ---
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
