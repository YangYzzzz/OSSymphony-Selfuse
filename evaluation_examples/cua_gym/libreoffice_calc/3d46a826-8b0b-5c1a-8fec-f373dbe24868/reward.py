"""
Reward Script: Calculate gross profit and write summary to Sheet2
Task ID: osworld_calc_gross_profit_sheet2_concat_015
Domain: libreoffice_calc
Scoring:
  Component 1: Column F has 'Total Cost' header + B+C+D formulas in F2:F13 (0.30 pts)
  Component 2: Column G has 'Gross Profit' header + E-F formulas in G2:G13 (0.30 pts)
  Component 3: Summary sheet A1 has INDEX/MATCH formula for best margin product (0.40 pts)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_gross_profit_sheet2_concat_015'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Verify the Manufacturing sheet and Summary sheet exist
    if 'Manufacturing' not in wb.sheetnames:
        print("CRITICAL: 'Manufacturing' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    if 'Summary' not in wb.sheetnames:
        print("CRITICAL: 'Summary' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws_mfg = wb['Manufacturing']
    ws_sum = wb['Summary']

    # =========================================================================
    # Component 1: Column F header = "Total Cost" AND F2:F13 contain =B+C+D formulas (0.30 pts)
    # In initial_env: columns F and G are empty (max_col=5)
    # In golden_env: F1="Total Cost", F2:F13 have formula =B+C+D pattern
    # =========================================================================
    try:
        f_header = ws_mfg.cell(row=1, column=6).value
        f_header_ok = (f_header is not None and
                       str(f_header).strip().lower() == 'total cost')

        # Check that F2:F13 all have formulas matching =B?+C?+D? pattern
        f_formulas_ok = True
        f_formula_count = 0
        for r in range(2, 14):  # rows 2-13 (12 products)
            cell_val = ws_mfg.cell(row=r, column=6).value
            if cell_val is None:
                f_formulas_ok = False
                print(f"FAIL: Component 1 — F{r} is empty, expected =B+C+D formula")
                break
            cell_str = str(cell_val).strip()
            # Accept formula like =B2+C2+D2 or similar total-cost pattern
            if isinstance(cell_val, str) and re.match(
                    r'=\s*B\d+\s*\+\s*C\d+\s*\+\s*D\d+', cell_str, re.IGNORECASE):
                f_formula_count += 1
            else:
                f_formulas_ok = False
                print(f"FAIL: Component 1 — F{r} value '{cell_val}' is not a =B+C+D formula")
                break

        if f_header_ok and f_formulas_ok and f_formula_count == 12:
            print(f"PASS: Component 1 — Column F header='{f_header}', all 12 rows have =B+C+D formulas (0.30 pts)")
            total_score += 0.30
        else:
            if not f_header_ok:
                print(f"FAIL: Component 1 — F1 header is '{f_header}', expected 'Total Cost'")
            elif not f_formulas_ok:
                pass  # Already printed failure detail above
            else:
                print(f"FAIL: Component 1 — Only {f_formula_count}/12 rows have valid formulas")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Column G header = "Gross Profit" AND G2:G13 contain =E-F formulas (0.30 pts)
    # In initial_env: column G is empty
    # In golden_env: G1="Gross Profit", G2:G13 have formula =E-F pattern
    # =========================================================================
    try:
        g_header = ws_mfg.cell(row=1, column=7).value
        g_header_ok = (g_header is not None and
                       str(g_header).strip().lower() == 'gross profit')

        # Check that G2:G13 all have formulas matching =E?-F? pattern
        g_formulas_ok = True
        g_formula_count = 0
        for r in range(2, 14):  # rows 2-13 (12 products)
            cell_val = ws_mfg.cell(row=r, column=7).value
            if cell_val is None:
                g_formulas_ok = False
                print(f"FAIL: Component 2 — G{r} is empty, expected =E-F formula")
                break
            cell_str = str(cell_val).strip()
            # Accept formula like =E2-F2 or similar gross-profit pattern
            if isinstance(cell_val, str) and re.match(
                    r'=\s*E\d+\s*-\s*F\d+', cell_str, re.IGNORECASE):
                g_formula_count += 1
            else:
                g_formulas_ok = False
                print(f"FAIL: Component 2 — G{r} value '{cell_val}' is not a =E-F formula")
                break

        if g_header_ok and g_formulas_ok and g_formula_count == 12:
            print(f"PASS: Component 2 — Column G header='{g_header}', all 12 rows have =E-F formulas (0.30 pts)")
            total_score += 0.30
        else:
            if not g_header_ok:
                print(f"FAIL: Component 2 — G1 header is '{g_header}', expected 'Gross Profit'")
            elif not g_formulas_ok:
                pass  # Already printed failure detail above
            else:
                print(f"FAIL: Component 2 — Only {g_formula_count}/12 rows have valid formulas")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Summary sheet A1 contains INDEX/MATCH formula for best margin product (0.40 pts)
    # In initial_env: Summary A1 is empty
    # In golden_env: Summary A1 has a formula string including INDEX, MATCH, and "Best Margin Product"
    # The formula must:
    #   (a) Be a formula string (starts with '=')
    #   (b) Reference INDEX and MATCH (for looking up the max gross profit product)
    #   (c) Contain "Best Margin Product" text in the concatenation
    #   (d) Format to 2 decimal places (use TEXT(...,"0.00") or similar)
    # =========================================================================
    try:
        a1_val = ws_sum.cell(row=1, column=1).value
        if a1_val is None:
            print("FAIL: Component 3 — Summary!A1 is empty")
        else:
            a1_str = str(a1_val).strip()
            is_formula = a1_str.startswith('=')
            has_index = 'INDEX' in a1_str.upper()
            has_match = 'MATCH' in a1_str.upper()
            has_best_margin = 'Best Margin Product' in a1_str
            # Check for 2 decimal places formatting (TEXT function with 0.00 or similar)
            has_decimal_format = ('0.00' in a1_str or '"0.00"' in a1_str)

            if is_formula and has_index and has_match and has_best_margin and has_decimal_format:
                print(f"PASS: Component 3 — Summary A1 has INDEX/MATCH formula with 'Best Margin Product' text and 2 decimal formatting (0.40 pts)")
                total_score += 0.40
            else:
                details = []
                if not is_formula:
                    details.append("not a formula (doesn't start with '=')")
                if not has_index:
                    details.append("missing INDEX function")
                if not has_match:
                    details.append("missing MATCH function")
                if not has_best_margin:
                    details.append("missing 'Best Margin Product' text")
                if not has_decimal_format:
                    details.append("missing 2-decimal formatting ('0.00')")
                print(f"FAIL: Component 3 — Summary A1 issues: {'; '.join(details)}")
                print(f"  Actual A1 value: {a1_str[:200]}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
