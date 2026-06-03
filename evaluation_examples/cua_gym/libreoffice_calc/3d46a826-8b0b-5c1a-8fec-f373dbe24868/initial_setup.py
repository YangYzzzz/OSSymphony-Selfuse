"""
Initial Setup: Calculate gross profit per unit and write summary in Sheet2
Task ID: osworld_calc_gross_profit_sheet2_concat_015
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_gross_profit_sheet2_concat_015'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Manufacturing ---
    ws1 = wb.active
    ws1.title = 'Manufacturing'

    # Headers in row 1 (columns A-E only; F and G are task targets, left empty)
    headers = ['Product', 'Material Cost', 'Labor Cost', 'Overhead', 'Selling Price']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # Realistic manufacturing data (12 rows)
    data = [
        ['Precision Gear Set',    42.50,  18.75,  9.30,   89.99],
        ['Hydraulic Valve Kit',   67.20,  24.10,  12.85, 139.95],
        ['Composite Frame Panel', 53.80,  31.40,  15.20, 125.00],
        ['Electronic Control Board', 89.40, 22.60, 18.75, 179.90],
        ['Stainless Steel Bracket', 14.20, 8.50,  4.10,   38.75],
        ['Polymer Seal Assembly',  22.60, 11.30,  5.85,   55.50],
        ['Titanium Fastener Pack', 38.90, 13.25,  7.40,   82.00],
        ['Aluminum Heat Sink',     27.15, 16.80,  8.60,   71.25],
        ['Rubber Vibration Damper', 9.80,  6.40,  3.20,   26.50],
        ['Carbon Fiber Rod',       74.50, 28.90,  14.30, 148.75],
        ['Copper Winding Coil',    31.60, 19.45,  9.70,   75.80],
        ['Ceramic Insulator Disc', 18.30, 10.75,  5.40,   43.20],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # --- Sheet 2: Summary ---
    # Sheet2 A1 is intentionally empty — task requires agent to add the summary formula
    ws2 = wb.create_sheet('Summary')
    ws2.cell(row=1, column=1, value=None)  # explicitly empty

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
