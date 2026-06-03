"""
Reward Script: Use VLOOKUP with approximate match to categorize employee satisfaction scores
Task ID: osworld_calc_vlookup_grade_lookup_004
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): VLOOKUP formula exists in column C data rows (C2:C16) with correct lookup range
  Component 2 (0.3): Formulas use approximate match (TRUE or 1) as the 4th argument
  Component 3 (0.3): All 15 data rows (C2:C16) have VLOOKUP formulas (complete coverage)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_vlookup_grade_lookup_004'

# Expected lookup table range (columns D-E, rows 2-5)
EXPECTED_LOOKUP_RANGE_PATTERNS = [
    r'\$?D\$?2:\$?E\$?5',   # D2:E5 or $D$2:$E$5 or $D2:$E5 etc.
    r'\$?D\$?2:\$?E\$?[5-9]',  # allow slightly larger range too
]

TOTAL_DATA_ROWS = 15  # rows 2 through 16


def has_vlookup_formula(cell_value):
    """Check if a cell value is a VLOOKUP formula string."""
    if not isinstance(cell_value, str):
        return False
    return cell_value.upper().strip().startswith('=VLOOKUP(')


def extract_vlookup_last_arg(formula):
    """
    Extract the last argument (range_lookup) from a VLOOKUP formula string.
    Returns the string after the last comma, stripped.
    VLOOKUP(lookup_value, table_array, col_index_num, [range_lookup])
    """
    if not isinstance(formula, str):
        return None
    # Remove leading = and whitespace
    f = formula.strip()
    if f.startswith('='):
        f = f[1:]
    # Simple extraction: find last comma before the closing paren
    # Account for nested parens
    depth = 0
    last_comma_pos = -1
    for i, ch in enumerate(f):
        if ch == '(':
            depth += 1
        elif ch == ')':
            depth -= 1
        elif ch == ',' and depth == 1:
            last_comma_pos = i
    if last_comma_pos == -1:
        return None
    # Extract from last comma to closing paren
    closing_paren = f.rfind(')')
    if closing_paren == -1:
        return None
    last_arg = f[last_comma_pos + 1:closing_paren].strip()
    return last_arg


def is_approximate_match(last_arg):
    """Check if the range_lookup argument indicates approximate match (TRUE or 1)."""
    if last_arg is None:
        return False
    normalized = last_arg.upper().strip()
    return normalized in ('TRUE', '1')


def check_lookup_range_reference(formula):
    """Check if the formula references the expected lookup table columns D-E rows 2-5."""
    if not isinstance(formula, str):
        return False
    # Check for D...E...2...5 pattern in the table_array argument
    for pattern in EXPECTED_LOOKUP_RANGE_PATTERNS:
        if re.search(pattern, formula, re.IGNORECASE):
            return True
    return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task: Column C (Category) in 'Survey Results' sheet should have VLOOKUP formulas
    using approximate match (TRUE) referencing the category table in D2:E5.
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet must exist
    if 'Survey Results' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Survey Results' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Survey Results']

    # Precondition: column C header must be 'Category'
    header_c = ws['C1'].value
    if not header_c or str(header_c).strip() != 'Category':
        print(f"CRITICAL: C1 header expected 'Category', found: {header_c}")
        print("REWARD: 0.0")
        return 0.0

    # --- Gather all column C data formulas (rows 2 to 16) ---
    c_formulas = []
    for row_idx in range(2, TOTAL_DATA_ROWS + 2):  # rows 2..16
        cell_val = ws.cell(row=row_idx, column=3).value
        c_formulas.append((row_idx, cell_val))

    vlookup_rows = [(r, v) for r, v in c_formulas if has_vlookup_formula(v)]
    non_empty_c = [(r, v) for r, v in c_formulas if v is not None]

    # Component 1: VLOOKUP formula with correct lookup range exists in column C (0.4 points)
    # This checks that at least some VLOOKUP formulas reference the correct D2:E5 table
    try:
        vlookup_with_correct_range = [
            (r, v) for r, v in vlookup_rows if check_lookup_range_reference(v)
        ]
        if len(vlookup_with_correct_range) > 0:
            print(f"PASS: Component 1 — {len(vlookup_with_correct_range)} VLOOKUP formula(s) found referencing D:E lookup table (0.4 pts)")
            total_score += 0.4
        else:
            if len(vlookup_rows) > 0:
                print(f"FAIL: Component 1 — {len(vlookup_rows)} VLOOKUP formula(s) found but none reference expected lookup table ($D$2:$E$5). Found: {vlookup_rows[0][1]}")
            elif len(non_empty_c) > 0:
                print(f"FAIL: Component 1 — Column C has values but no VLOOKUP formulas. Found: {non_empty_c[0][1]}")
            else:
                print("FAIL: Component 1 — No VLOOKUP formulas in column C (C2:C16 all empty)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: VLOOKUP formulas use approximate match (TRUE/1 as 4th argument) (0.3 points)
    # This checks the critical requirement: approximate match for categorization by thresholds
    try:
        if len(vlookup_with_correct_range) > 0:
            approx_match_rows = [
                (r, v) for r, v in vlookup_with_correct_range
                if is_approximate_match(extract_vlookup_last_arg(v))
            ]
            if len(approx_match_rows) > 0:
                # Check that ALL formulas with correct range use approximate match
                all_approx = (len(approx_match_rows) == len(vlookup_with_correct_range))
                if all_approx:
                    print(f"PASS: Component 2 — All {len(approx_match_rows)} VLOOKUP formulas use approximate match (TRUE) (0.3 pts)")
                    total_score += 0.3
                else:
                    # Partial: some but not all use approximate match — still award if majority do
                    ratio = len(approx_match_rows) / len(vlookup_with_correct_range)
                    if ratio >= 0.5:
                        print(f"PASS: Component 2 — {len(approx_match_rows)}/{len(vlookup_with_correct_range)} VLOOKUP formulas use approximate match (TRUE) (0.3 pts)")
                        total_score += 0.3
                    else:
                        print(f"FAIL: Component 2 — Only {len(approx_match_rows)}/{len(vlookup_with_correct_range)} formulas use approximate match (TRUE). Check 4th argument.")
            else:
                sample = vlookup_with_correct_range[0][1]
                last_arg = extract_vlookup_last_arg(sample)
                print(f"FAIL: Component 2 — VLOOKUP formulas do NOT use approximate match. 4th arg found: '{last_arg}' (expected TRUE or 1). Formula sample: {sample}")
        else:
            print("SKIP: Component 2 — No qualifying VLOOKUP formulas to check for approximate match")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Complete coverage — all 15 data rows (C2:C16) have VLOOKUP formulas (0.3 points)
    # This checks that the task was applied to every employee row, not just some
    try:
        covered_count = len(vlookup_rows)
        if covered_count == TOTAL_DATA_ROWS:
            print(f"PASS: Component 3 — All {TOTAL_DATA_ROWS} data rows (C2:C16) have VLOOKUP formulas (0.3 pts)")
            total_score += 0.3
        elif covered_count > 0:
            print(f"FAIL: Component 3 — Only {covered_count}/{TOTAL_DATA_ROWS} rows have VLOOKUP formulas. Rows without: {[r for r, v in c_formulas if not has_vlookup_formula(v)]}")
        else:
            print(f"FAIL: Component 3 — No VLOOKUP formulas found in C2:C16 (0/{TOTAL_DATA_ROWS} covered)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
