"""
Initial Setup: Employee satisfaction survey with VLOOKUP grade lookup task
Task ID: osworld_calc_vlookup_grade_lookup_004
Domain: libreoffice_calc

Creates a spreadsheet with:
- Employee ID and satisfaction score data (columns A-B)
- Empty Category column C (to be filled by agent with VLOOKUP)
- Reference table with score thresholds and category labels (columns D-E)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_vlookup_grade_lookup_004'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Survey Results ---
    ws = wb.active
    ws.title = 'Survey Results'

    # Headers for data columns A-C
    header_font = Font(bold=True)
    ws['A1'] = 'Employee ID'
    ws['B1'] = 'Score'
    ws['C1'] = 'Category'
    ws['A1'].font = header_font
    ws['B1'].font = header_font
    ws['C1'].font = header_font

    # Reference table headers in D-E
    ws['D1'] = 'Min Score'
    ws['E1'] = 'Category Label'
    ws['D1'].font = header_font
    ws['E1'].font = header_font

    # Employee survey data - realistic employee IDs and varied scores
    # Scores deliberately cover all 4 tiers: Disengaged(<40), Neutral(40-59), Satisfied(60-79), Highly Satisfied(80+)
    employee_data = [
        ('EMP-1042', 73),
        ('EMP-2317', 88),
        ('EMP-0854', 22),
        ('EMP-3601', 55),
        ('EMP-1789', 91),
        ('EMP-2456', 45),
        ('EMP-0923', 67),
        ('EMP-3182', 14),
        ('EMP-1504', 82),
        ('EMP-2891', 38),
        ('EMP-0765', 61),
        ('EMP-3347', 95),
        ('EMP-1628', 49),
        ('EMP-2073', 77),
        ('EMP-0391', 33),
    ]

    for r, (emp_id, score) in enumerate(employee_data, 2):
        ws.cell(row=r, column=1, value=emp_id)
        ws.cell(row=r, column=2, value=score)
        # Column C (Category) is intentionally left EMPTY - agent must fill with VLOOKUP

    # Reference table: score thresholds and category labels (columns D-E)
    # Sorted ascending (required for VLOOKUP approximate match)
    category_table = [
        (0,  'Disengaged'),
        (40, 'Neutral'),
        (60, 'Satisfied'),
        (80, 'Highly Satisfied'),
    ]
    for r, (min_score, label) in enumerate(category_table, 2):
        ws.cell(row=r, column=4, value=min_score)
        ws.cell(row=r, column=5, value=label)

    # Column width adjustments for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
