"""
Reward Script: Clean up order database by removing duplicate orders
Task ID: calc_dop_dedup_multi_036
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Result has exactly 182 data rows (18 duplicates removed from 200)
  Component 2 (0.40): All rows are unique by (Customer ID, Product Code, Order Date) — no duplicates remain
  Component 3 (0.30): Kept rows are the FIRST occurrences (duplicate Order Refs from ORD-20001..ORD-20018 are absent)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_dop_dedup_multi_036'

# These are the Order Ref values of the 18 duplicate rows that should be REMOVED.
# They are the second occurrences of each duplicated (Customer ID, Product Code, Order Date) key.
# Identified from the initial file: when the same key appears twice, the second
# occurrence has Order Ref in the ORD-20001..ORD-20018 range.
DUPLICATE_ORDER_REFS = {
    'ORD-20001', 'ORD-20002', 'ORD-20003', 'ORD-20004', 'ORD-20005',
    'ORD-20006', 'ORD-20007', 'ORD-20008', 'ORD-20009', 'ORD-20010',
    'ORD-20011', 'ORD-20012', 'ORD-20013', 'ORD-20014', 'ORD-20015',
    'ORD-20016', 'ORD-20017', 'ORD-20018',
}

EXPECTED_DATA_ROWS = 182  # 200 original - 18 duplicates removed


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: 'Orders' sheet must exist
    if 'Orders' not in wb.sheetnames:
        print("CRITICAL: 'Orders' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Orders']

    # Precondition gate: headers must be intact
    expected_headers = ['Customer ID', 'Order Ref', 'Product Code', 'Quantity', 'Order Date', 'Amount']
    actual_headers = [ws.cell(row=1, column=col).value for col in range(1, 7)]
    if actual_headers != expected_headers:
        print(f"CRITICAL: Headers corrupted. Expected {expected_headers}, found {actual_headers}")
        print("REWARD: 0.0")
        return 0.0

    actual_data_rows = ws.max_row - 1  # subtract header row

    # Component 1: Exactly 182 data rows remain after deduplication (0.30 points)
    # Initial file has 200 rows; this FAILS on initial (200 != 182)
    try:
        if actual_data_rows == EXPECTED_DATA_ROWS:
            print(f"PASS: Component 1 — Correct row count: {actual_data_rows} data rows ({EXPECTED_DATA_ROWS} expected) (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 — Expected {EXPECTED_DATA_ROWS} data rows, found {actual_data_rows}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: No duplicate (Customer ID, Product Code, Order Date) keys remain (0.40 points)
    # Initial file has 18 duplicates; this FAILS on initial
    try:
        seen_keys = {}
        duplicate_count = 0
        for row in range(2, ws.max_row + 1):
            cust = ws.cell(row=row, column=1).value
            prod = ws.cell(row=row, column=3).value
            date = ws.cell(row=row, column=5).value
            key = (cust, prod, date)
            if key in seen_keys:
                duplicate_count += 1
            else:
                seen_keys[key] = row

        if duplicate_count == 0:
            print(f"PASS: Component 2 — No duplicate (Customer ID, Product Code, Order Date) keys found (0.40 pts)")
            total_score += 0.40
        else:
            print(f"FAIL: Component 2 — Found {duplicate_count} duplicate key(s) still present")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Duplicate second-occurrence rows are absent (first occurrences were kept) (0.30 points)
    # The 18 duplicate entries have Order Refs ORD-20001..ORD-20018 (the second occurrences).
    # If any of these are still in the file, the wrong rows were kept or duplicates were not removed.
    # This FAILS on initial because all 18 ORD-20001..ORD-20018 refs are present there.
    try:
        found_dup_refs = []
        for row in range(2, ws.max_row + 1):
            order_ref = ws.cell(row=row, column=2).value
            if order_ref in DUPLICATE_ORDER_REFS:
                found_dup_refs.append(order_ref)

        if len(found_dup_refs) == 0:
            print(f"PASS: Component 3 — All 18 duplicate second-occurrence Order Refs (ORD-20001..ORD-20018) correctly removed (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 3 — Found {len(found_dup_refs)} duplicate-occurrence Order Ref(s) still present: {found_dup_refs[:5]}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
