"""
Initial Setup: Create my_recipes.xlsx with tried recipes list
Task ID: osworld_multi_apps_misc_024
Domain: libreoffice_calc (multi-app: LibreOffice Calc + Chrome)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_misc_024'
OUTPUT = f'{WORKDIR}/my_recipes.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: my_recipes ---
    ws = wb.active
    ws.title = 'my_recipes'

    # Headers
    headers = ['Rating', 'Recipe Name', 'Category', 'Prep Time']
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF', size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Tried recipes - realistic data. These are popular recipes the user has already tried.
    # Some overlap with AllRecipes top-15, some don't.
    tried_data = [
        [4.8, "Best Chocolate Chip Cookies", "Desserts", "15 mins"],
        [4.9, "Simple Roast Chicken", "Meat and Poultry", "20 mins"],
        [4.7, "Classic Beef Lasagna", "Pasta", "45 mins"],
        [4.6, "Homemade Banana Bread", "Bread", "15 mins"],
        [4.5, "Creamy Tomato Basil Soup", "Soups, Stews and Chili", "20 mins"],
        [4.8, "Beef and Broccoli Stir Fry", "World Cuisine", "15 mins"],
        [4.4, "Blueberry Muffins", "Bread", "15 mins"],
        [4.7, "Chicken Tikka Masala", "World Cuisine", "30 mins"],
        [4.6, "Classic Caesar Salad", "Salad", "15 mins"],
        [4.5, "Homemade Beef Tacos", "Main Dishes", "15 mins"],
        [4.8, "Fluffy Buttermilk Pancakes", "Breakfast and Brunch", "10 mins"],
        [4.3, "Vegetable Minestrone Soup", "Soups, Stews and Chili", "25 mins"],
        [4.7, "Garlic Butter Shrimp Pasta", "Pasta", "20 mins"],
        [4.6, "Classic Beef Pot Roast", "Meat and Poultry", "30 mins"],
        [4.5, "Chocolate Lava Cake", "Desserts", "15 mins"],
    ]

    for r, row_data in enumerate(tried_data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open Chrome with AllRecipes, then open LibreOffice Calc
    launch_gui('google-chrome "https://www.allrecipes.com"', delay_sec=3.0)
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched Chrome with AllRecipes and LibreOffice Calc with DISPLAY=:0')


create_initial()
