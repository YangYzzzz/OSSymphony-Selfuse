"""
Reward Script: Add 'recipes_to_try' sheet to my_recipes.xlsx with AllRecipes top-15
Task ID: osworld_multi_apps_misc_024
Domain: libreoffice_calc
Scoring:
  Component 1: 'recipes_to_try' sheet exists in workbook (0.30)
  Component 2: Headers in 'recipes_to_try' match required columns (0.20)
  Component 3: Data rows are sorted by Rating descending (0.30)
  Component 4: No recipe in 'recipes_to_try' is already in 'my_recipes' (0.20)
  Total: 1.00
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_misc_024'

REQUIRED_HEADERS = ['Rating', 'Recipe Name', 'Category', 'Prep Time']


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook — critical gate
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get existing recipes from the 'my_recipes' sheet (precondition: must exist)
    tried_recipes = set()
    if 'my_recipes' in wb.sheetnames:
        ws_tried = wb['my_recipes']
        for row in ws_tried.iter_rows(min_row=2, max_row=ws_tried.max_row, values_only=True):
            if row[1]:  # Recipe Name is column B (index 1)
                tried_recipes.add(str(row[1]).strip().lower())

    # Component 1: 'recipes_to_try' sheet exists (0.30 points)
    try:
        if 'recipes_to_try' in wb.sheetnames:
            print("PASS: Component 1 — 'recipes_to_try' sheet exists (0.30 pts)")
            total_score += 0.30
            ws_new = wb['recipes_to_try']
        else:
            print(f"FAIL: Component 1 — 'recipes_to_try' sheet not found. Sheets: {wb.sheetnames}")
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Component 2: Correct headers in 'recipes_to_try' (0.20 points)
    try:
        actual_headers = [ws_new.cell(row=1, column=c).value for c in range(1, ws_new.max_column + 1)]
        # Normalize to strip whitespace
        actual_headers_stripped = [str(h).strip() if h is not None else '' for h in actual_headers[:4]]
        if actual_headers_stripped == REQUIRED_HEADERS:
            print(f"PASS: Component 2 — Headers match {REQUIRED_HEADERS} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 — Expected headers {REQUIRED_HEADERS}, found {actual_headers_stripped}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Data rows exist and are sorted by Rating descending (0.30 points)
    try:
        data_rows = list(ws_new.iter_rows(min_row=2, max_row=ws_new.max_row, values_only=True))
        # Filter out completely empty rows
        data_rows = [r for r in data_rows if any(v is not None for v in r)]
        if len(data_rows) == 0:
            print("FAIL: Component 3 — No data rows in 'recipes_to_try'")
        else:
            ratings = []
            invalid_rating = None
            for row in data_rows:
                try:
                    rating = float(row[0])
                    ratings.append(rating)
                except (TypeError, ValueError):
                    invalid_rating = row[0]
                    break

            if invalid_rating is not None:
                print(f"FAIL: Component 3 — Non-numeric rating found: {invalid_rating}")
            elif len(ratings) > 0:
                is_sorted_desc = ratings == sorted(ratings, reverse=True)
                if is_sorted_desc:
                    print(f"PASS: Component 3 — {len(data_rows)} rows sorted by Rating descending (0.30 pts). First: {ratings[0]}, Last: {ratings[-1]}")
                    total_score += 0.30
                else:
                    print(f"FAIL: Component 3 — Ratings not sorted descending. Found order: {ratings}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: No recipe in 'recipes_to_try' overlaps with 'my_recipes' (0.20 points)
    try:
        new_recipe_rows = list(ws_new.iter_rows(min_row=2, max_row=ws_new.max_row, values_only=True))
        new_recipe_rows = [r for r in new_recipe_rows if any(v is not None for v in r)]
        if len(new_recipe_rows) == 0:
            print("FAIL: Component 4 — No data rows to check for duplicates")
        else:
            overlap = []
            for row in new_recipe_rows:
                if row[1] and str(row[1]).strip().lower() in tried_recipes:
                    overlap.append(str(row[1]).strip())
            if len(overlap) == 0:
                print(f"PASS: Component 4 — No overlap between 'recipes_to_try' and 'my_recipes' (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 4 — Found {len(overlap)} recipe(s) already in 'my_recipes': {overlap}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the VM environment
file_path = f'{WORKDIR}/my_recipes.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
