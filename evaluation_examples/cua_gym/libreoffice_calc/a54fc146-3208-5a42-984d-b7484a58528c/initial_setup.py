"""
Initial Setup: Add data validation to project dates in LibreOffice Calc
Task ID: calc_ggf_020
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
from datetime import date, timedelta
import random

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_020'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Projects'

    # --- Headers ---
    headers = ['Project ID', 'Name', 'Lead', 'Due Date', 'Status']
    header_font = Font(bold=True, size=11, name='Calibri')
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, name='Calibri', color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # --- Column widths ---
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14

    # --- Realistic project data (30 rows) ---
    project_names = [
        'Cloud Migration Phase 2', 'CRM System Upgrade', 'Mobile App Redesign',
        'Data Warehouse Optimization', 'Security Audit Framework', 'API Gateway Deployment',
        'Customer Portal v3', 'Inventory Management Overhaul', 'HR Onboarding Automation',
        'Payment Processing Integration', 'Supply Chain Analytics', 'DevOps Pipeline Setup',
        'Compliance Reporting Tool', 'Marketing Dashboard', 'Employee Training Platform',
        'Vendor Management System', 'Disaster Recovery Plan', 'Network Infrastructure Upgrade',
        'Quality Assurance Automation', 'Business Intelligence Suite', 'Warehouse Robotics Pilot',
        'Client Feedback System', 'Document Management Portal', 'Financial Forecasting Engine',
        'Product Recommendation AI', 'Sustainability Tracking App', 'Fleet Management System',
        'Digital Signature Integration', 'Customer Loyalty Program', 'Real-Time Monitoring Dashboard',
    ]

    leads = [
        'Sarah Chen', 'Marcus Johnson', 'Emily Rodriguez', 'David Kim',
        'Priya Patel', 'James O\'Brien', 'Lisa Nakamura', 'Ahmed Hassan',
        'Rachel Goldstein', 'Carlos Mendez', 'Sofia Andersson', 'Wei Zhang',
        'Olivia Thompson', 'Michael Brown', 'Aisha Williams', 'Thomas Mueller',
        'Jennifer Park', 'Robert Singh', 'Anna Kowalski', 'Daniel Lee',
        'Maya Gupta', 'Kevin O\'Connor', 'Fatima Al-Rashid', 'Nathan Brooks',
        'Isabella Cruz', 'Victor Petrov', 'Hannah Schmidt', 'Oscar Reyes',
        'Grace Liu', 'Patrick Doyle',
    ]

    statuses = ['In Progress', 'Planning', 'On Hold', 'In Review', 'Active', 'Scoping']

    # Generate dates - most within 2024-2025 but some deliberately outside
    random.seed(42)
    dates = []
    for i in range(30):
        if i == 3:
            # Deliberately out of range: too old
            dates.append(date(2020, 6, 15))
        elif i == 12:
            # Deliberately out of range: too old
            dates.append(date(2020, 11, 3))
        elif i == 21:
            # Deliberately out of range: too far future
            dates.append(date(2030, 3, 22))
        elif i == 27:
            # Deliberately out of range: too far future
            dates.append(date(2030, 8, 10))
        else:
            # Valid date within 2024-2025
            start_d = date(2024, 1, 1)
            end_d = date(2025, 12, 31)
            delta = (end_d - start_d).days
            rand_days = random.randint(0, delta)
            dates.append(start_d + timedelta(days=rand_days))

    for r in range(30):
        row_num = r + 2
        ws.cell(row=row_num, column=1, value=f'PRJ-{1001 + r}')
        ws.cell(row=row_num, column=2, value=project_names[r])
        ws.cell(row=row_num, column=3, value=leads[r])

        date_cell = ws.cell(row=row_num, column=4, value=dates[r])
        date_cell.number_format = 'yyyy-mm-dd'

        ws.cell(row=row_num, column=5, value=statuses[r % len(statuses)])

    # Apply light borders to data rows
    for r in range(2, 32):
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.border = header_border

    # NO data validation on D2:D31 - that's the task!

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
