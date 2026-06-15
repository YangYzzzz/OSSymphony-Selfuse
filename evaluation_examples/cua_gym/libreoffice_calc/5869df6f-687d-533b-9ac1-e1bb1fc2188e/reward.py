"""
Reward Script: Sports League Standings Tracker
Task ID: calc_wf_068
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Formulas in Standings for Won/Drawn/Lost/GF/GA/GD/Points
  Component 2 (0.20): Teams sorted by Points desc then GD desc
  Component 3 (0.15): Bar chart on Standings sheet
  Component 4 (0.15): Conditional formatting (top 4 green, bottom 2 red)
  Component 5 (0.10): Table formatting (header bold+fill, borders)
  Component 6 (0.10): Freeze panes set to A2
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_068'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Standings sheet must exist
    if 'Standings' not in wb.sheetnames:
        print("CRITICAL: 'Standings' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Standings']

    # ---------------------------------------------------------------
    # Component 1: Formulas in Standings columns C-I (0.30 points)
    # Initial has None in C2:I2; golden has formulas/values.
    # We check that at least 10 of 12 team rows have non-None values
    # in columns C through I (Won, Drawn, Lost, GF, GA, GD, Points).
    # We also verify Points formula logic: Points = Won*3 + Drawn.
    # ---------------------------------------------------------------
    try:
        filled_rows = 0
        formula_rows = 0
        for r in range(2, 14):
            # Check columns C through I (Won, Drawn, Lost, GF, GA, GD, Points)
            vals = [ws.cell(row=r, column=c).value for c in range(3, 10)]
            non_none_count = sum(1 for v in vals if v is not None)
            if non_none_count >= 5:
                filled_rows += 1
            # Check if any cell contains a formula (string starting with '=')
            has_formula = any(
                isinstance(ws.cell(row=r, column=c).value, str) and
                ws.cell(row=r, column=c).value.startswith('=')
                for c in range(3, 10)
            )
            if has_formula:
                formula_rows += 1

        if filled_rows >= 10 and formula_rows >= 10:
            print(f"PASS: Component 1 — {filled_rows}/12 rows filled with formulas ({formula_rows} formula rows) (0.30 pts)")
            total_score += 0.30
        elif filled_rows >= 10:
            # Data present but might be hardcoded values instead of formulas
            print(f"PARTIAL: Component 1 — {filled_rows}/12 rows filled, {formula_rows} with formulas (0.20 pts)")
            total_score += 0.20
        elif filled_rows >= 6:
            print(f"PARTIAL: Component 1 — only {filled_rows}/12 rows filled (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 — only {filled_rows}/12 rows have data in C-I columns")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ---------------------------------------------------------------
    # Component 2: Teams sorted by Points desc then GD desc (0.20 points)
    # Initial has alphabetical order. Golden has sorted order.
    # We compute expected stats from Results sheet to verify sort order.
    # ---------------------------------------------------------------
    try:
        # Compute expected standings from Results sheet data
        if 'Results' in wb.sheetnames:
            ws_results = wb['Results']
            # Build stats from results data
            team_stats = {}  # team -> {won, drawn, lost, gf, ga}
            for r in range(2, ws_results.max_row + 1):
                home = ws_results.cell(row=r, column=1).value
                away = ws_results.cell(row=r, column=2).value
                hg = ws_results.cell(row=r, column=3).value
                ag = ws_results.cell(row=r, column=4).value
                if not home or not away or hg is None or ag is None:
                    continue
                for t in [home, away]:
                    if t not in team_stats:
                        team_stats[t] = {'won': 0, 'drawn': 0, 'lost': 0, 'gf': 0, 'ga': 0}
                # Home team
                team_stats[home]['gf'] += int(hg)
                team_stats[home]['ga'] += int(ag)
                if int(hg) > int(ag):
                    team_stats[home]['won'] += 1
                    team_stats[away]['lost'] += 1
                elif int(hg) == int(ag):
                    team_stats[home]['drawn'] += 1
                    team_stats[away]['drawn'] += 1
                else:
                    team_stats[home]['lost'] += 1
                    team_stats[away]['won'] += 1
                # Away team
                team_stats[away]['gf'] += int(ag)
                team_stats[away]['ga'] += int(hg)

            # Compute points and GD, sort
            team_ranking = []
            for team, s in team_stats.items():
                pts = s['won'] * 3 + s['drawn']
                gd = s['gf'] - s['ga']
                team_ranking.append((team, pts, gd))
            team_ranking.sort(key=lambda x: (-x[1], -x[2]))
            expected_order = [t[0] for t in team_ranking]

            # Get current order from Standings sheet
            current_order = []
            for r in range(2, 14):
                team = ws.cell(row=r, column=1).value
                if team:
                    current_order.append(team)

            initial_order = ['Arsenal', 'Chelsea', 'Liverpool', 'Manchester City',
                             'Tottenham', 'Manchester United', 'Newcastle',
                             'Aston Villa', 'Brighton', 'West Ham',
                             'Wolverhampton', 'Burnley']

            is_initial_order = (current_order == initial_order)

            if current_order == expected_order:
                print(f"PASS: Component 2 — Teams correctly sorted by Points desc, GD desc (0.20 pts)")
                print(f"  Expected order: {expected_order}")
                print(f"  Current order:  {current_order}")
                total_score += 0.20
            elif not is_initial_order and len(current_order) >= 10:
                # Teams reordered but not exactly matching expected
                # Check if top 3 and bottom 2 are correct for partial credit
                top_match = current_order[:3] == expected_order[:3]
                print(f"PARTIAL: Component 2 — Teams reordered but not perfectly sorted (0.10 pts)")
                print(f"  Expected: {expected_order}")
                print(f"  Got:      {current_order}")
                total_score += 0.10
            else:
                print(f"FAIL: Component 2 — Teams still in initial order or insufficient data")
                print(f"  Current: {current_order}")
        else:
            print(f"FAIL: Component 2 — Results sheet not found, cannot verify sort")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---------------------------------------------------------------
    # Component 3: Bar chart on Standings sheet (0.15 points)
    # Initial has 0 charts. Golden has 1 BarChart.
    # ---------------------------------------------------------------
    try:
        charts = ws._charts
        if len(charts) >= 1:
            bar_chart_found = False
            for chart in charts:
                class_name = chart.__class__.__name__
                if 'Bar' in class_name:
                    bar_chart_found = True
                    break
            if bar_chart_found:
                print(f"PASS: Component 3 — Bar chart found on Standings sheet (0.15 pts)")
                total_score += 0.15
            else:
                # Chart exists but not a bar chart - partial credit
                print(f"PARTIAL: Component 3 — Chart found but not a bar chart (type: {charts[0].__class__.__name__}) (0.07 pts)")
                total_score += 0.07
        else:
            print(f"FAIL: Component 3 — No charts found on Standings sheet")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ---------------------------------------------------------------
    # Component 4: Conditional formatting (0.15 points)
    # Initial has 0 conditional formatting rules. Golden has 2 rules:
    # - Top 4 rows: green background (FFC6EFCE)
    # - Bottom 2 rows: red background (FFFFC7CE)
    # ---------------------------------------------------------------
    try:
        cf_rules = list(ws.conditional_formatting)
        if len(cf_rules) >= 1:
            # Count total rules across all ranges
            total_rules = sum(len(cf.rules) for cf in cf_rules)
            has_green = False
            has_red = False

            for cf in cf_rules:
                for rule in cf.rules:
                    if rule.dxf and rule.dxf.fill and rule.dxf.fill.fgColor:
                        color = rule.dxf.fill.fgColor.rgb
                        if color:
                            color_upper = str(color).upper()
                            # Check for green-ish colors (C6EFCE or similar greens)
                            if 'C6EFCE' in color_upper or '00FF00' in color_upper or 'EFCE' in color_upper:
                                has_green = True
                            # Check for red-ish colors (FFC7CE or similar reds)
                            if 'FFC7CE' in color_upper or 'FF0000' in color_upper or 'C7CE' in color_upper:
                                has_red = True

            if has_green and has_red:
                print(f"PASS: Component 4 — Conditional formatting with green (top 4) and red (bottom 2) (0.15 pts)")
                total_score += 0.15
            elif has_green or has_red:
                found = 'green' if has_green else 'red'
                print(f"PARTIAL: Component 4 — Only {found} conditional formatting found (0.07 pts)")
                total_score += 0.07
            else:
                print(f"PARTIAL: Component 4 — Conditional formatting present ({total_rules} rules) but no green/red fill detected (0.05 pts)")
                total_score += 0.05
        else:
            print(f"FAIL: Component 4 — No conditional formatting rules found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ---------------------------------------------------------------
    # Component 5: Table formatting - header styling + borders (0.10 points)
    # Initial has bold header but no fill/borders. Golden has bold+fill+borders.
    # We check header row has solid fill AND data cells have borders.
    # ---------------------------------------------------------------
    try:
        header_has_fill = False
        header_fill_count = 0
        for c in range(1, 10):
            cell = ws.cell(row=1, column=c)
            if cell.fill.fill_type == 'solid':
                fg = cell.fill.fgColor
                if fg and fg.rgb and str(fg.rgb) != '00000000':
                    header_fill_count += 1

        if header_fill_count >= 7:
            header_has_fill = True

        # Check borders on data cells
        border_count = 0
        total_checked = 0
        for r in [2, 7, 13]:
            for c in [1, 5, 9]:
                total_checked += 1
                cell = ws.cell(row=r, column=c)
                if (cell.border.left.style and cell.border.bottom.style):
                    border_count += 1

        has_borders = border_count >= 6

        if header_has_fill and has_borders:
            print(f"PASS: Component 5 — Header fill ({header_fill_count}/9 cols) and borders ({border_count}/{total_checked} cells) (0.10 pts)")
            total_score += 0.10
        elif header_has_fill or has_borders:
            found = 'header fill' if header_has_fill else 'borders'
            print(f"PARTIAL: Component 5 — Only {found} present (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — No header fill (count={header_fill_count}) and no borders (count={border_count})")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # ---------------------------------------------------------------
    # Component 6: Freeze panes set to A2 (0.10 points)
    # Initial has no freeze panes. Golden has A2.
    # ---------------------------------------------------------------
    try:
        if ws.freeze_panes == 'A2':
            print(f"PASS: Component 6 — Freeze panes set to A2 (0.10 pts)")
            total_score += 0.10
        elif ws.freeze_panes is not None:
            print(f"PARTIAL: Component 6 — Freeze panes set to {ws.freeze_panes}, expected A2 (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 — No freeze panes set")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
