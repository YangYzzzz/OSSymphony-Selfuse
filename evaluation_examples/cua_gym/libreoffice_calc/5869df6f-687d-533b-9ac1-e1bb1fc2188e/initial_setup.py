"""
Initial Setup: Sports league standings tracker
Task ID: calc_wf_068
Domain: libreoffice_calc

Creates a workbook with:
- 'Results' sheet: 30 match results for 12 teams
- 'Standings' sheet: 12 teams with headers but empty computed columns
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_068'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

TEAMS = [
    'Arsenal', 'Chelsea', 'Liverpool', 'Manchester City',
    'Tottenham', 'Manchester United', 'Newcastle', 'Aston Villa',
    'Brighton', 'West Ham', 'Wolverhampton', 'Burnley'
]

# Match results: (home_team, away_team, home_goals, away_goals)
MATCHES = [
    ('Arsenal', 'Chelsea', 2, 1),
    ('Liverpool', 'Manchester City', 3, 2),
    ('Tottenham', 'Manchester United', 1, 1),
    ('Newcastle', 'Aston Villa', 2, 0),
    ('Brighton', 'West Ham', 1, 2),
    ('Wolverhampton', 'Burnley', 3, 1),
    ('Chelsea', 'Liverpool', 0, 2),
    ('Manchester City', 'Arsenal', 1, 1),
    ('Manchester United', 'Newcastle', 3, 2),
    ('Aston Villa', 'Tottenham', 0, 1),
    ('West Ham', 'Wolverhampton', 2, 2),
    ('Burnley', 'Brighton', 1, 3),
    ('Arsenal', 'Liverpool', 1, 0),
    ('Chelsea', 'Manchester City', 2, 2),
    ('Tottenham', 'Newcastle', 4, 1),
    ('Manchester United', 'Brighton', 2, 0),
    ('Aston Villa', 'West Ham', 3, 1),
    ('Wolverhampton', 'Arsenal', 0, 2),
    ('Burnley', 'Chelsea', 1, 4),
    ('Liverpool', 'Tottenham', 2, 1),
    ('Manchester City', 'Newcastle', 3, 0),
    ('Brighton', 'Aston Villa', 1, 1),
    ('West Ham', 'Manchester United', 0, 1),
    ('Arsenal', 'Tottenham', 3, 1),
    ('Chelsea', 'Newcastle', 1, 0),
    ('Liverpool', 'Burnley', 5, 0),
    ('Manchester City', 'West Ham', 4, 1),
    ('Manchester United', 'Wolverhampton', 2, 1),
    ('Aston Villa', 'Burnley', 2, 0),
    ('Brighton', 'Wolverhampton', 2, 1),
]


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Results ---
    ws_results = wb.active
    ws_results.title = 'Results'

    # Headers
    headers = ['Home Team', 'Away Team', 'Home Goals', 'Away Goals']
    for col, h in enumerate(headers, 1):
        cell = ws_results.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11)

    # Match data
    for r, match in enumerate(MATCHES, 2):
        for c, val in enumerate(match, 1):
            ws_results.cell(row=r, column=c, value=val)

    # Column widths
    ws_results.column_dimensions['A'].width = 22
    ws_results.column_dimensions['B'].width = 22
    ws_results.column_dimensions['C'].width = 14
    ws_results.column_dimensions['D'].width = 14

    # --- Sheet 2: Standings ---
    ws_standings = wb.create_sheet('Standings')

    standing_headers = ['Team', 'Played', 'Won', 'Drawn', 'Lost',
                        'GF', 'GA', 'GD', 'Points']
    for col, h in enumerate(standing_headers, 1):
        cell = ws_standings.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11)

    # Pre-compute how many games each team played (for the Played column only)
    from collections import Counter
    games_played = Counter()
    for home, away, hg, ag in MATCHES:
        games_played[home] += 1
        games_played[away] += 1

    # Fill team names and Played column; leave all other columns EMPTY
    for r, team in enumerate(TEAMS, 2):
        ws_standings.cell(row=r, column=1, value=team)
        ws_standings.cell(row=r, column=2, value=games_played.get(team, 0))

    # Column widths
    ws_standings.column_dimensions['A'].width = 22
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws_standings.column_dimensions[col_letter].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
