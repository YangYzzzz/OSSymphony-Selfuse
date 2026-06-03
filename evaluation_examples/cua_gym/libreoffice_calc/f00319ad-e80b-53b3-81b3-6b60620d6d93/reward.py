"""
Reward Script: Apply accounting number format to B2:B4
Task ID: calc_lf_079
Domain: libreoffice_calc
Scoring:
  Component 1: B2 has accounting format (0.35 pts)
  Component 2: B3 has accounting format (0.35 pts)
  Component 3: B4 has accounting format (0.30 pts)
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_079'

# The expected accounting format string (exact match from golden env).
# We also accept close variants that are functionally equivalent.
EXPECTED_FORMAT = '_("$"* #,##0.00_);_("$"* (#,##0.00);_("$"* "-"??_)'


def is_accounting_format(fmt_str):
    """
    Check if the number format is an accounting-style format.
    Must contain the key structural elements:
    - Uses _( for alignment padding
    - Contains "$" currency symbol
    - Uses #,##0.00 number pattern
    - Has separate sections for positive, negative, and zero values
    Rejects: 'General', plain currency like '$#,##0.00', or empty formats.
    """
    if not fmt_str or fmt_str == 'General':
        return False

    # Normalize whitespace for comparison
    normalized = fmt_str.replace(' ', '')

    # Key markers of the accounting format:
    # 1. Has alignment padding _( or _)
    has_padding = '_(' in fmt_str or '_)' in fmt_str
    # 2. Contains dollar sign reference
    has_dollar = '$' in fmt_str
    # 3. Contains the number pattern #,##0.00
    has_number_pattern = '#,##0.00' in fmt_str
    # 4. Has multiple sections (semicolons for pos;neg;zero)
    has_sections = fmt_str.count(';') >= 1

    return has_padding and has_dollar and has_number_pattern and has_sections


def persist_app_state(domain):
    """Save any unsaved LibreOffice edits via Ctrl+S."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet exists
    if 'Ledger' not in wb.sheetnames:
        print("FAIL: 'Ledger' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Ledger']

    # Component 1: B2 has accounting format (0.35 points)
    try:
        fmt_b2 = ws['B2'].number_format
        if is_accounting_format(fmt_b2):
            print(f"PASS: Component 1 -- B2 has accounting format: {fmt_b2} (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 1 -- B2 format is '{fmt_b2}', expected accounting format")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: B3 has accounting format (0.35 points)
    try:
        fmt_b3 = ws['B3'].number_format
        if is_accounting_format(fmt_b3):
            print(f"PASS: Component 2 -- B3 has accounting format: {fmt_b3} (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 2 -- B3 format is '{fmt_b3}', expected accounting format")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: B4 has accounting format (0.30 points)
    try:
        fmt_b4 = ws['B4'].number_format
        if is_accounting_format(fmt_b4):
            print(f"PASS: Component 3 -- B4 has accounting format: {fmt_b4} (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 3 -- B4 format is '{fmt_b4}', expected accounting format")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state('libreoffice_calc')

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
