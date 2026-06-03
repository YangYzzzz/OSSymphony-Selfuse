"""
Initial Setup: Set up data validation on D2:D25 for future dates
Task ID: calc_gcv_074
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_074'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reservation_System"

    # --- Headers ---
    headers = ["Reservation ID", "Guest Name", "Room Type", "Check-in Date"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Data: 24 reservations (rows 2-25) ---
    reservations = [
        ["RSV-1001", "Sarah Chen", "Deluxe Suite"],
        ["RSV-1002", "Marcus Johnson", "Standard Double"],
        ["RSV-1003", "Elena Vasquez", "Executive Suite"],
        ["RSV-1004", "Raj Patel", "Standard Single"],
        ["RSV-1005", "Olivia Thompson", "Deluxe Double"],
        ["RSV-1006", "Kenji Tanaka", "Penthouse Suite"],
        ["RSV-1007", "Amara Okafor", "Standard Double"],
        ["RSV-1008", "Lucas Bergmann", "Deluxe Suite"],
        ["RSV-1009", "Sofia Morales", "Executive Suite"],
        ["RSV-1010", "David Kim", "Standard Single"],
        ["RSV-1011", "Isabella Rossi", "Deluxe Double"],
        ["RSV-1012", "James O'Brien", "Standard Double"],
        ["RSV-1013", "Fatima Al-Hassan", "Penthouse Suite"],
        ["RSV-1014", "Henrik Larsson", "Deluxe Suite"],
        ["RSV-1015", "Priya Sharma", "Executive Suite"],
        ["RSV-1016", "Thomas Mitchell", "Standard Single"],
        ["RSV-1017", "Yuki Watanabe", "Deluxe Double"],
        ["RSV-1018", "Anna Kowalski", "Standard Double"],
        ["RSV-1019", "Carlos Rivera", "Executive Suite"],
        ["RSV-1020", "Mei Lin Zhang", "Deluxe Suite"],
        ["RSV-1021", "Benjamin Foster", "Standard Double"],
        ["RSV-1022", "Nadia Petrova", "Penthouse Suite"],
        ["RSV-1023", "Samuel Adeyemi", "Deluxe Double"],
        ["RSV-1024", "Claire Dubois", "Standard Single"],
    ]

    data_align = Alignment(horizontal="left", vertical="center")
    date_format = 'yyyy-mm-dd'

    for r, row_data in enumerate(reservations, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = data_align
            cell.border = thin_border
        # Column D: empty but formatted as date
        d_cell = ws.cell(row=r, column=4)
        d_cell.number_format = date_format
        d_cell.alignment = data_align
        d_cell.border = thin_border

    # --- Column widths ---
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 18

    # --- Freeze header row ---
    ws.freeze_panes = "A2"

    # NO data validation on the initial file
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
