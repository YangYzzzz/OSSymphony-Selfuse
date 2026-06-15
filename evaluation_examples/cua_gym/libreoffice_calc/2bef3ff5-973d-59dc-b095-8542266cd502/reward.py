"""
Reward Script: Remove open password from SecureData.xlsx
Task ID: calc_ps_045
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): File is a valid ZIP (not OLE2 encrypted)
  Component 2 (0.3): File loads in openpyxl without password, both sheets exist
  Component 3 (0.3): Data integrity - spot-check key values across both sheets
"""

import os
import zipfile

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_045'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Component 1: File is a valid ZIP container, not OLE2 encrypted (0.4 points)
    # An encrypted xlsx is stored as OLE2 compound document (header D0CF11E0).
    # A password-free xlsx is a standard ZIP container (header 504B0304).
    try:
        with open(file_path, 'rb') as f:
            header = f.read(4)

        if header == b'\x50\x4b\x03\x04':
            # Confirm it's actually a valid zip
            try:
                zf = zipfile.ZipFile(file_path)
                zf.close()
                print(f"PASS: Component 1 — File is a valid ZIP (not encrypted) (0.4 pts)")
                total_score += 0.4
            except zipfile.BadZipFile:
                print(f"FAIL: Component 1 — File has ZIP header but is not a valid ZIP")
        elif header == b'\xd0\xcf\x11\xe0':
            print(f"FAIL: Component 1 — File is OLE2 compound document (still encrypted)")
        else:
            print(f"FAIL: Component 1 — Unknown file format, header: {header.hex()}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: File loads in openpyxl and has expected sheets (0.3 points)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
        sheets = wb.sheetnames

        if 'Financial Summary' in sheets and 'Employee Records' in sheets:
            print(f"PASS: Component 2 — File loads without password, both sheets present: {sheets} (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — Expected sheets 'Financial Summary' and 'Employee Records', found: {sheets}")
        wb.close()
    except Exception as e:
        print(f"FAIL: Component 2 — Cannot load file with openpyxl (likely still encrypted): {e}")

    # Component 3: Data integrity - spot-check key values across both sheets (0.3 points)
    # This ensures the password was removed without corrupting data.
    # Combined with Component 1 check: only awards points if file is unencrypted AND data intact.
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)

        integrity_checks_passed = 0
        total_integrity_checks = 4

        # Check Financial Summary sheet
        ws_fin = wb['Financial Summary']

        # Check header row
        if ws_fin['A1'].value == 'Quarter' and ws_fin['C1'].value == 'Revenue':
            integrity_checks_passed += 1
        else:
            print(f"  INFO: Financial Summary headers mismatch: A1={ws_fin['A1'].value}, C1={ws_fin['C1'].value}")

        # Check a specific data value: C2 should be 245300 (Q1 Cloud Services Revenue)
        if ws_fin['C2'].value == 245300:
            integrity_checks_passed += 1
        else:
            print(f"  INFO: Financial Summary C2 expected 245300, found: {ws_fin['C2'].value}")

        # Check Employee Records sheet
        ws_emp = wb['Employee Records']

        # Check header
        if ws_emp['A1'].value == 'Employee ID' and ws_emp['B1'].value == 'Full Name':
            integrity_checks_passed += 1
        else:
            print(f"  INFO: Employee Records headers mismatch: A1={ws_emp['A1'].value}, B1={ws_emp['B1'].value}")

        # Check a specific employee: B2 should be 'Sarah Chen'
        if ws_emp['B2'].value == 'Sarah Chen':
            integrity_checks_passed += 1
        else:
            print(f"  INFO: Employee Records B2 expected 'Sarah Chen', found: {ws_emp['B2'].value}")

        if integrity_checks_passed == total_integrity_checks:
            print(f"PASS: Component 3 — Data integrity verified ({integrity_checks_passed}/{total_integrity_checks} checks) (0.3 pts)")
            total_score += 0.3
        else:
            # Partial credit within this component
            partial = round(0.3 * (integrity_checks_passed / total_integrity_checks), 2)
            print(f"PARTIAL: Component 3 — {integrity_checks_passed}/{total_integrity_checks} integrity checks passed ({partial} pts)")
            total_score += partial

        wb.close()
    except Exception as e:
        print(f"FAIL: Component 3 — Cannot verify data integrity (likely still encrypted): {e}")

    final_score = round(min(total_score, 1.0), 1)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/SecureData.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
