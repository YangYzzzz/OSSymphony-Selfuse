"""
Initial Setup: Create a password-protected workbook SecureData.xlsx
Task ID: calc_ps_045
Domain: libreoffice_calc

The workbook has an open password 'open123' and contains data across two sheets:
- Sheet 'Financial Summary' with quarterly revenue data
- Sheet 'Employee Records' with employee information
"""

import io
import os
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from msoffcrypto.method.ecma376_agile import ECMA376Agile

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_045'
FILENAME = 'SecureData.xlsx'
OUTPUT = f'{WORKDIR}/{FILENAME}'
PASSWORD = 'open123'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_workbook():
    """Create the workbook content (shared structure for initial and golden)."""
    wb = openpyxl.Workbook()

    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    # --- Sheet 1: Financial Summary ---
    ws1 = wb.active
    ws1.title = 'Financial Summary'

    headers1 = ['Quarter', 'Product Line', 'Revenue', 'Cost of Goods', 'Gross Profit', 'Margin %']
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    data1 = [
        ['Q1 2025', 'Cloud Services', 245300, 98120, 147180, 0.60],
        ['Q1 2025', 'Hardware', 182750, 127925, 54825, 0.30],
        ['Q1 2025', 'Consulting', 96400, 57840, 38560, 0.40],
        ['Q2 2025', 'Cloud Services', 268900, 107560, 161340, 0.60],
        ['Q2 2025', 'Hardware', 175200, 122640, 52560, 0.30],
        ['Q2 2025', 'Consulting', 112800, 67680, 45120, 0.40],
        ['Q3 2025', 'Cloud Services', 291450, 116580, 174870, 0.60],
        ['Q3 2025', 'Hardware', 163800, 114660, 49140, 0.30],
        ['Q3 2025', 'Consulting', 105600, 63360, 42240, 0.40],
        ['Q4 2025', 'Cloud Services', 312800, 125120, 187680, 0.60],
        ['Q4 2025', 'Hardware', 198500, 138950, 59550, 0.30],
        ['Q4 2025', 'Consulting', 128900, 77340, 51560, 0.40],
    ]

    for r, row_data in enumerate(data1, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c in (3, 4, 5):
                cell.number_format = '$#,##0'
            elif c == 6:
                cell.number_format = '0%'

    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 14
    ws1.column_dimensions['D'].width = 16
    ws1.column_dimensions['E'].width = 14
    ws1.column_dimensions['F'].width = 12

    # --- Sheet 2: Employee Records ---
    ws2 = wb.create_sheet('Employee Records')

    headers2 = ['Employee ID', 'Full Name', 'Department', 'Position', 'Annual Salary', 'Hire Date', 'Status']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    data2 = [
        ['EMP-001', 'Sarah Chen', 'Engineering', 'Senior Developer', 125000, '2021-03-15', 'Active'],
        ['EMP-002', 'Marcus Johnson', 'Marketing', 'Campaign Manager', 89000, '2022-06-01', 'Active'],
        ['EMP-003', 'Priya Patel', 'Finance', 'Financial Analyst', 95000, '2020-11-20', 'Active'],
        ["EMP-004", "James O'Brien", 'Engineering', 'DevOps Engineer', 118000, '2023-01-10', 'Active'],
        ['EMP-005', 'Ana Rodriguez', 'HR', 'HR Specialist', 78000, '2021-08-05', 'Active'],
        ['EMP-006', 'Wei Zhang', 'Engineering', 'QA Lead', 105000, '2019-04-22', 'Active'],
        ['EMP-007', 'Emily Watson', 'Sales', 'Account Executive', 92000, '2022-09-14', 'Active'],
        ['EMP-008', 'David Kim', 'Finance', 'Controller', 135000, '2018-02-28', 'Active'],
        ['EMP-009', 'Lisa Thompson', 'Marketing', 'Content Strategist', 82000, '2023-05-16', 'Active'],
        ['EMP-010', 'Robert Garcia', 'Engineering', 'Tech Lead', 142000, '2020-07-01', 'Active'],
        ['EMP-011', 'Michelle Lee', 'Sales', 'Sales Director', 155000, '2017-11-30', 'Active'],
        ['EMP-012', 'Tom Anderson', 'HR', 'Recruiter', 72000, '2024-01-08', 'Active'],
    ]

    for r, row_data in enumerate(data2, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 5:
                cell.number_format = '$#,##0'

    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 16
    ws2.column_dimensions['D'].width = 20
    ws2.column_dimensions['E'].width = 16
    ws2.column_dimensions['F'].width = 14
    ws2.column_dimensions['G'].width = 10

    return wb


def create_initial():
    wb = create_workbook()

    # Save unencrypted workbook to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Encrypt with open password using ECMA-376 Agile Encryption
    encrypted_bytes = ECMA376Agile.encrypt(key=PASSWORD, ibuf=buf)

    # Write encrypted file to disk
    with open(OUTPUT, 'wb') as f:
        f.write(encrypted_bytes)

    print(f'Initial encrypted file created: {OUTPUT}')

    # Launch LibreOffice Calc (agent will need to enter password to open)
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
