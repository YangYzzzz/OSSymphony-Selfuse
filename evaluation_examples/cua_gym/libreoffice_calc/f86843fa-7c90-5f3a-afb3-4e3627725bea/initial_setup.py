"""
Initial Setup: Apply data validation to cell range F2:F100
Task ID: calc_gao_038
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gao_038'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Products'

    # --- Headers ---
    headers = ['Product ID', 'Product Name', 'Category', 'SKU', 'Cost Price', 'Selling Price']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Column widths ---
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 16

    # --- Product data (columns A-E populated, F is EMPTY) ---
    products = [
        ['P-1001', 'Wireless Bluetooth Headphones', 'Electronics', 'SKU-WBH-4421', 24.50],
        ['P-1002', 'Organic Green Tea (100 bags)', 'Beverages', 'SKU-OGT-8837', 8.75],
        ['P-1003', 'Stainless Steel Water Bottle', 'Kitchenware', 'SKU-SSW-2219', 12.30],
        ['P-1004', 'USB-C Charging Cable 2m', 'Electronics', 'SKU-UCC-5564', 5.20],
        ['P-1005', 'Bamboo Cutting Board Set', 'Kitchenware', 'SKU-BCB-1198', 18.90],
        ['P-1006', 'Premium Dark Roast Coffee', 'Beverages', 'SKU-PDR-3356', 14.60],
        ['P-1007', 'LED Desk Lamp Adjustable', 'Electronics', 'SKU-LDL-7742', 32.40],
        ['P-1008', 'Cotton Bath Towel Set (4)', 'Home & Living', 'SKU-CBT-9901', 22.15],
        ['P-1009', 'Yoga Mat Non-Slip 6mm', 'Sports', 'SKU-YMN-6623', 16.80],
        ['P-1010', 'Ceramic Dinner Plate Set', 'Kitchenware', 'SKU-CDP-4457', 28.50],
        ['P-1011', 'Portable Power Bank 10000mAh', 'Electronics', 'SKU-PPB-8814', 19.75],
        ['P-1012', 'Lavender Essential Oil 30ml', 'Health & Beauty', 'SKU-LEO-2238', 9.40],
        ['P-1013', 'Hardcover Notebook A5', 'Stationery', 'SKU-HNA-5571', 6.25],
        ['P-1014', 'Silicone Baking Mat Set', 'Kitchenware', 'SKU-SBM-3394', 11.50],
        ['P-1015', 'Wireless Mouse Ergonomic', 'Electronics', 'SKU-WME-7718', 15.90],
        ['P-1016', 'Reusable Shopping Bags (5)', 'Home & Living', 'SKU-RSB-1142', 7.80],
        ['P-1017', 'Stainless Steel Lunch Box', 'Kitchenware', 'SKU-SSL-6689', 21.30],
        ['P-1018', 'Resistance Band Set (5 pcs)', 'Sports', 'SKU-RBS-4456', 13.25],
        ['P-1019', 'Scented Soy Candle 200g', 'Home & Living', 'SKU-SSC-8872', 10.60],
        ['P-1020', 'Mechanical Pencil Set', 'Stationery', 'SKU-MPS-2215', 4.85],
        ['P-1021', 'Glass Food Storage Containers', 'Kitchenware', 'SKU-GFS-9938', 25.70],
        ['P-1022', 'Bluetooth Speaker Mini', 'Electronics', 'SKU-BSM-5563', 29.90],
        ['P-1023', 'Herbal Chamomile Tea (50 bags)', 'Beverages', 'SKU-HCT-1197', 6.50],
        ['P-1024', 'Microfiber Cleaning Cloths (10)', 'Home & Living', 'SKU-MCC-7741', 8.20],
        ['P-1025', 'Adjustable Dumbbell 5-25kg', 'Sports', 'SKU-ADB-3384', 45.00],
    ]

    data_font = Font(name='Calibri', size=11)
    currency_format = '$#,##0.00'

    for r, row_data in enumerate(products, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = thin_border
            if c == 5:  # Cost Price column
                cell.number_format = currency_format
                cell.alignment = Alignment(horizontal='right')

    # F column cells (F2:F26) - leave empty but apply currency format for readability
    for r in range(2, len(products) + 2):
        cell = ws.cell(row=r, column=6)
        cell.number_format = currency_format
        cell.border = thin_border

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
