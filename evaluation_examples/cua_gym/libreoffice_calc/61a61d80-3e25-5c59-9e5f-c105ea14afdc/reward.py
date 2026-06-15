"""
Reward Script: Turn off gridlines and row/column headers in 'Slide Template' sheet
Task ID: calc_sht_gridlines_002
Domain: libreoffice_calc
Scoring:
  Component 1: 'Slide Template' sheet has gridlines hidden (showGridLines=False) — 0.5 pts
  Component 2: 'Slide Template' sheet has row/column headers hidden (showRowColHeaders=False) — 0.5 pts
  Note: 'Data Source' sheet changes are precondition gates (must remain unchanged), not scored separately.
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_sht_gridlines_002'

SLIDE_TEMPLATE_SHEET = 'Slide Template'
DATA_SOURCE_SHEET = 'Data Source'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Task: Turn off gridlines and row/column headers in the 'Slide Template' sheet.
    The 'Data Source' sheet must remain unmodified.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook — if this fails, we cannot proceed
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: verify both expected sheets exist
    if SLIDE_TEMPLATE_SHEET not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SLIDE_TEMPLATE_SHEET}' not found in workbook. "
              f"Available sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    if DATA_SOURCE_SHEET not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{DATA_SOURCE_SHEET}' not found in workbook. "
              f"Available sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: verify 'Data Source' sheet gridlines/headers were NOT modified
    # (Task says 'Data Source' must remain with gridlines and headers visible)
    try:
        ws_data = wb[DATA_SOURCE_SHEET]
        sv_data = ws_data.sheet_view
        if sv_data.showGridLines is False:
            print(f"FAIL (gate): 'Data Source' sheet has gridlines hidden — should remain visible. "
                  f"Data integrity check failed.")
            print("REWARD: 0.0")
            return 0.0
        if sv_data.showRowColHeaders is False:
            print(f"FAIL (gate): 'Data Source' sheet has row/col headers hidden — should remain visible. "
                  f"Data integrity check failed.")
            print("REWARD: 0.0")
            return 0.0
        print(f"PASS (gate): 'Data Source' sheet retains visible gridlines and headers as expected.")
    except Exception as e:
        print(f"ERROR (gate): Could not check 'Data Source' sheet view: {e}")
        # Not blocking — continue scoring

    ws = wb[SLIDE_TEMPLATE_SHEET]

    # Component 1: 'Slide Template' gridlines are hidden (showGridLines=False) — 0.5 points
    # This FAILS on initial (showGridLines=True) and PASSES on golden (showGridLines=False)
    try:
        sv = ws.sheet_view
        show_gridlines = sv.showGridLines
        if show_gridlines is False:
            print(f"PASS: Component 1 — 'Slide Template' gridlines are hidden "
                  f"(showGridLines=False) (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — 'Slide Template' gridlines still visible "
                  f"(showGridLines={show_gridlines}), expected False")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not check showGridLines: {e}")

    # Component 2: 'Slide Template' row/column headers are hidden (showRowColHeaders=False) — 0.5 points
    # This FAILS on initial (showRowColHeaders=True) and PASSES on golden (showRowColHeaders=False)
    try:
        sv = ws.sheet_view
        show_headers = sv.showRowColHeaders
        if show_headers is False:
            print(f"PASS: Component 2 — 'Slide Template' row/column headers are hidden "
                  f"(showRowColHeaders=False) (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 2 — 'Slide Template' row/column headers still visible "
                  f"(showRowColHeaders={show_headers}), expected False")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not check showRowColHeaders: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
