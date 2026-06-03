"""
Initial Setup: Turn off gridlines and row/column headers in the 'Slide Template' sheet
Task ID: calc_sht_gridlines_002
Domain: libreoffice_calc

Creates a workbook with two sheets:
  - 'Slide Template': Styled like a PowerPoint slide with logo area, title, and body.
    Gridlines and row/column headers are VISIBLE (pre-task state).
  - 'Data Source': Raw data with gridlines and headers visible.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WORKDIR = '/home/user'
TASK_ID = 'calc_sht_gridlines_002'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # ----------------------------------------------------------------
    # Sheet 1: Slide Template
    # ----------------------------------------------------------------
    ws1 = wb.active
    ws1.title = 'Slide Template'

    # Ensure gridlines AND row/col headers are visible (default state)
    ws1.sheet_view.showGridLines = True
    ws1.sheet_view.showRowColHeaders = True

    # --- Set column widths for a slide-like layout ---
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws1.column_dimensions[col_letter].width = 14

    # --- Merged cells A1:H3 — Company logo area ---
    ws1.merge_cells('A1:H3')
    logo_cell = ws1['A1']
    logo_cell.value = 'Nexus Analytics Corp.'
    logo_cell.font = Font(name='Calibri', size=22, bold=True, color='FFFFFF')
    logo_cell.fill = PatternFill(start_color='FF1F3864', end_color='FF1F3864', fill_type='solid')
    logo_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws1.row_dimensions[1].height = 20
    ws1.row_dimensions[2].height = 20
    ws1.row_dimensions[3].height = 20

    # --- A4:H4 — Slide Title ---
    ws1.merge_cells('A4:H4')
    title_cell = ws1['A4']
    title_cell.value = 'Q3 2025 Business Performance Review'
    title_cell.font = Font(name='Calibri', size=20, bold=True, color='FF1F3864')
    title_cell.fill = PatternFill(start_color='FFD6E4F0', end_color='FFD6E4F0', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[4].height = 36

    # --- Row 5: spacer ---
    ws1.row_dimensions[5].height = 10

    # --- A6:H25 — Body content area ---
    # Section header
    ws1.merge_cells('A6:H6')
    section_cell = ws1['A6']
    section_cell.value = 'Executive Summary'
    section_cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFFFF')
    section_cell.fill = PatternFill(start_color='FF2E75B6', end_color='FF2E75B6', fill_type='solid')
    section_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws1.row_dimensions[6].height = 24

    # Body bullet points (A7:H25)
    bullets = [
        ('A7:H7',  'Total Revenue exceeded $4.2M, up 18% from Q2 2025'),
        ('A8:H8',  'Operating Margin improved to 23.4%, driven by cost efficiencies'),
        ('A9:H9',  'New customer acquisitions: 342 accounts (target: 300)'),
        ('A10:H10', 'Customer retention rate: 94.7% — highest in company history'),
        ('A11:H11', ''),
        ('A12:H12', 'Key Achievements'),
        ('A13:H13', '  • Launched the Nexus Cloud Platform v3.0 in August'),
        ('A14:H14', '  • Expanded to 3 new regional markets: Denver, Miami, Portland'),
        ('A15:H15', '  • Reduced average support ticket resolution time by 31%'),
        ('A16:H16', '  • Completed ISO 27001 security certification'),
        ('A17:H17', ''),
        ('A18:H18', 'Challenges & Mitigation'),
        ('A19:H19', '  • Supply chain delays affected hardware delivery timelines'),
        ('A20:H20', '  • Mitigation: Secured alternate vendor contracts with 2-week SLA'),
        ('A21:H21', '  • Talent acquisition slower than planned (72% of Q3 hiring goal)'),
        ('A22:H22', '  • Mitigation: Partnered with 4 university placement programs'),
        ('A23:H23', ''),
        ('A24:H24', 'Outlook: Q4 2025 revenue target set at $4.8M with 3 product launches'),
        ('A25:H25', 'Prepared by: Strategy & Finance Team  |  Confidential'),
    ]

    for merge_range, text in bullets:
        ws1.merge_cells(merge_range)
        start_coord = merge_range.split(':')[0]
        cell = ws1[start_coord]
        cell.value = text
        row_num = int(start_coord[1:])

        if text == 'Key Achievements' or text == 'Challenges & Mitigation':
            cell.font = Font(name='Calibri', size=12, bold=True, color='FF1F3864')
            cell.fill = PatternFill(start_color='FFE2EFDA', end_color='FFE2EFDA', fill_type='solid')
        elif text == '':
            pass
        elif 'Confidential' in text:
            cell.font = Font(name='Calibri', size=9, italic=True, color='FF808080')
            cell.alignment = Alignment(horizontal='right', vertical='center')
        else:
            cell.font = Font(name='Calibri', size=11, color='FF333333')
            cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)

        ws1.row_dimensions[row_num].height = 18

    # ----------------------------------------------------------------
    # Sheet 2: Data Source
    # ----------------------------------------------------------------
    ws2 = wb.create_sheet('Data Source')

    # Gridlines and row/col headers visible (default)
    ws2.sheet_view.showGridLines = True
    ws2.sheet_view.showRowColHeaders = True

    # --- Column widths ---
    col_widths = {'A': 20, 'B': 16, 'C': 16, 'D': 14, 'E': 14, 'F': 16}
    for col_letter, width in col_widths.items():
        ws2.column_dimensions[col_letter].width = width

    # --- Headers ---
    headers = ['Region', 'Product Line', 'Q3 Revenue ($)', 'Q3 Units Sold', 'Q2 Revenue ($)', 'YoY Growth (%)']
    thin = Side(style='thin', color='FF000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
        cell.fill = PatternFill(start_color='FF1F3864', end_color='FF1F3864', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = header_border
    ws2.row_dimensions[1].height = 30

    # --- Data rows ---
    data_rows = [
        ('North East',   'Analytics Suite',  892400,  1247, 756200,  18.0),
        ('North East',   'Cloud Platform',   534600,   823, 480100,  11.4),
        ('South West',   'Analytics Suite',  714300,  1089, 601500,  18.8),
        ('South West',   'Cloud Platform',   423700,   651, 378900,  11.8),
        ('Midwest',      'Analytics Suite',  638900,   952, 545600,  17.1),
        ('Midwest',      'Cloud Platform',   312400,   480, 275000,  13.6),
        ('Pacific',      'Analytics Suite',  521700,   789, 447800,  16.5),
        ('Pacific',      'Cloud Platform',   289300,   445, 256700,  12.7),
        ('South East',   'Analytics Suite',  463200,   698, 391400,  18.3),
        ('South East',   'Cloud Platform',   198600,   305, 174800,  13.6),
        ('International','Analytics Suite',  405800,   612, 340200,  19.3),
        ('International','Cloud Platform',   187300,   287, 155600,  20.4),
    ]

    alt_fill = PatternFill(start_color='FFF0F4F8', end_color='FFF0F4F8', fill_type='solid')
    normal_fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')

    for r, row_data in enumerate(data_rows, 2):
        fill = alt_fill if r % 2 == 0 else normal_fill
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.font = Font(name='Calibri', size=10)
            cell.border = header_border
            cell.fill = fill
            if c == 3 or c == 5:
                cell.number_format = '#,##0'
            elif c == 6:
                cell.number_format = '0.0'

    # --- Totals row ---
    total_row = len(data_rows) + 2
    ws2.cell(row=total_row, column=1, value='TOTAL').font = Font(name='Calibri', size=10, bold=True)
    ws2.cell(row=total_row, column=1).border = header_border
    ws2.cell(row=total_row, column=2, value='All Products').font = Font(name='Calibri', size=10, bold=True)
    ws2.cell(row=total_row, column=2).border = header_border

    for c in range(3, 7):
        cell = ws2.cell(row=total_row, column=c)
        col_letter = get_column_letter(c)
        cell.value = f'=SUM({col_letter}2:{col_letter}{total_row - 1})'
        cell.font = Font(name='Calibri', size=10, bold=True)
        cell.border = header_border
        cell.fill = PatternFill(start_color='FFD6E4F0', end_color='FFD6E4F0', fill_type='solid')
        if c == 3 or c == 5:
            cell.number_format = '#,##0'
        elif c == 6:
            cell.number_format = '0.0'

    # --- Freeze header row ---
    ws2.freeze_panes = 'A2'

    # ----------------------------------------------------------------
    # Save
    # ----------------------------------------------------------------
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: Slide Template (gridlines ON, headers ON), Data Source (gridlines ON, headers ON)')


create_initial()
