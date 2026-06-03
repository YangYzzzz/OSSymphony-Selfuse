"""
Initial Setup: Physics Data spreadsheet with wavelength values in General format
Task ID: calc_fmt_numfmt_scientific_025
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_numfmt_scientific_025'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Physics Data'

    # --- Headers ---
    ws['A1'] = 'Experiment'
    ws['B1'] = 'Wavelength (m)'
    ws['C1'] = 'Frequency (Hz)'
    ws['D1'] = 'Energy (J)'

    # Bold headers
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}1'].font = Font(bold=True)

    # --- Data rows 2-20 (19 experiments) ---
    # Wavelengths in meters (visible light and UV/IR range), stored as plain floats
    # Frequencies and energies calculated to be physically consistent
    # Using General format (default) for column B — no number_format set
    data = [
        ('Exp-001',  0.0000000450,  6.66e14,  4.41e-19),
        ('Exp-002',  0.0000000380,  7.89e14,  5.23e-19),
        ('Exp-003',  0.0000000620,  4.84e14,  3.21e-19),
        ('Exp-004',  0.0000000510,  5.88e14,  3.90e-19),
        ('Exp-005',  0.0000000730,  4.11e14,  2.72e-19),
        ('Exp-006',  0.0000000410,  7.32e14,  4.85e-19),
        ('Exp-007',  0.0000000560,  5.36e14,  3.55e-19),
        ('Exp-008',  0.0000000480,  6.25e14,  4.14e-19),
        ('Exp-009',  0.0000000390,  7.69e14,  5.09e-19),
        ('Exp-010',  0.0000000650,  4.62e14,  3.06e-19),
        ('Exp-011',  0.0000000425,  7.06e14,  4.67e-19),
        ('Exp-012',  0.0000000580,  5.17e14,  3.42e-19),
        ('Exp-013',  0.0000000350,  8.57e14,  5.67e-19),
        ('Exp-014',  0.0000000700,  4.28e14,  2.84e-19),
        ('Exp-015',  0.0000000465,  6.45e14,  4.27e-19),
        ('Exp-016',  0.0000000530,  5.66e14,  3.75e-19),
        ('Exp-017',  0.0000000395,  7.59e14,  5.02e-19),
        ('Exp-018',  0.0000000610,  4.92e14,  3.26e-19),
        ('Exp-019',  0.0000000440,  6.82e14,  4.51e-19),
    ]

    for r, (exp, wl, freq, energy) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=exp)
        ws.cell(row=r, column=2, value=wl)   # General format (no number_format set)
        ws.cell(row=r, column=3, value=freq)
        ws.cell(row=r, column=4, value=energy)

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet: Physics Data')
    print(f'Rows: 20 (1 header + 19 data rows)')
    print(f'Column B number format: General (task requires applying 0.00E+00)')


create_initial()
