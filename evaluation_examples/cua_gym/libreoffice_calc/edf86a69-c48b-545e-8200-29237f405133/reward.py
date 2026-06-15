"""
Reward Script: Apply scientific notation format (0.00E+00) to cells B2:B20
Task ID: calc_fmt_numfmt_scientific_025
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Fraction of cells B2:B20 that have format '0.00E+00' (0.7 points max)
  Component 2: ALL 19 cells B2:B20 have format '0.00E+00' (bonus 0.3 points for full completion)
  Total: 1.0

Note: Only format changes are scored. Values are precondition-checked (gate), not scored.
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_numfmt_scientific_025'

SHEET_NAME = 'Physics Data'
TARGET_FORMAT = '0.00E+00'
TARGET_COL = 2   # Column B
START_ROW = 2
END_ROW = 20
TOTAL_CELLS = END_ROW - START_ROW + 1  # 19

def verify_task(file_path):
    """
    Verify that B2:B20 have scientific notation format '0.00E+00'.
    Returns a float between 0.0 and 1.0.
    """
    total_score = 0.0

    # Gate: can we load the file?
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Gate: does the expected sheet exist?
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Gate: verify data integrity — underlying values in B2:B20 must be numeric
    # (This is a precondition, not scored. If data is corrupted, return 0.0 early.)
    try:
        for row in range(START_ROW, END_ROW + 1):
            val = ws.cell(row=row, column=TARGET_COL).value
            if val is None:
                print(f"CRITICAL: Cell B{row} is None — data integrity issue")
                print("REWARD: 0.0")
                return 0.0
            try:
                float(val)
            except (ValueError, TypeError):
                print(f"CRITICAL: Cell B{row} value '{val}' is not numeric — data integrity issue")
                print("REWARD: 0.0")
                return 0.0
    except Exception as e:
        print(f"CRITICAL: Data integrity check failed: {e}")
        print("REWARD: 0.0")
        return 0.0

    print(f"GATE PASS: Data integrity — all B2:B20 cells are numeric")

    # Component 1: Count how many cells in B2:B20 have format '0.00E+00' (0.7 points proportional)
    # This FAILS on initial (all cells have 'General') and PASSES on golden (all have '0.00E+00')
    cells_with_correct_format = 0
    cells_checked = []
    try:
        for row in range(START_ROW, END_ROW + 1):
            cell = ws.cell(row=row, column=TARGET_COL)
            fmt = cell.number_format
            if fmt == TARGET_FORMAT:
                cells_with_correct_format += 1
                cells_checked.append(f"B{row}:OK")
            else:
                cells_checked.append(f"B{row}:{fmt}")

        fraction = cells_with_correct_format / TOTAL_CELLS
        component1_score = round(fraction * 0.7, 4)

        print(f"Component 1: {cells_with_correct_format}/{TOTAL_CELLS} cells in B2:B20 have format '{TARGET_FORMAT}'")
        if cells_with_correct_format < TOTAL_CELLS:
            # Show first few failures
            failures = [c for c in cells_checked if c.endswith('General') or (not c.endswith(':OK'))]
            print(f"  Failures (first 5): {failures[:5]}")
        print(f"  Score: {component1_score:.4f}/0.7 ({cells_with_correct_format}/{TOTAL_CELLS} cells correct)")
        total_score += component1_score

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: ALL 19 cells have '0.00E+00' format (bonus 0.3 for complete coverage)
    # FAILS on initial (none have the format) and PASSES on golden (all 19 have it)
    try:
        if cells_with_correct_format == TOTAL_CELLS:
            print(f"PASS: Component 2 — All {TOTAL_CELLS} cells in B2:B20 have format '{TARGET_FORMAT}' (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — Only {cells_with_correct_format}/{TOTAL_CELLS} cells have correct format (need all {TOTAL_CELLS})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
