"""
Initial Setup: Insert a new sheet named 'Summary' after the last existing sheet
Task ID: calc_gsi_001
Domain: libreoffice_calc

Creates a workbook with three sheets: January, February, March.
Each sheet contains realistic monthly expense/revenue data.
No 'Summary' sheet exists yet - that is the agent's task.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_001'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Common styles ---
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    headers = ["Date", "Category", "Description", "Amount ($)", "Status"]

    # --- Sheet 1: January ---
    ws1 = wb.active
    ws1.title = "January"

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    jan_data = [
        ["2025-01-03", "Office Supplies", "Printer cartridges and paper", 234.50, "Approved"],
        ["2025-01-05", "Travel", "Flight to Seattle for client meeting", 487.00, "Approved"],
        ["2025-01-08", "Software", "Annual Figma license renewal", 1200.00, "Approved"],
        ["2025-01-10", "Catering", "Team lunch - project kickoff", 312.75, "Approved"],
        ["2025-01-14", "Equipment", "Ergonomic keyboard (x3)", 389.97, "Pending"],
        ["2025-01-17", "Marketing", "Google Ads campaign - Q1 launch", 2500.00, "Approved"],
        ["2025-01-20", "Training", "AWS certification course - Sarah Chen", 650.00, "Approved"],
        ["2025-01-22", "Utilities", "Internet service - January billing", 189.99, "Approved"],
        ["2025-01-25", "Office Supplies", "Whiteboard markers and sticky notes", 67.30, "Approved"],
        ["2025-01-28", "Travel", "Hotel stay - Marcus Johnson Chicago trip", 756.00, "Pending"],
        ["2025-01-30", "Software", "Slack workspace upgrade", 840.00, "Approved"],
        ["2025-01-31", "Maintenance", "HVAC system quarterly service", 425.00, "Approved"],
    ]

    for r, row_data in enumerate(jan_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 4:
                cell.number_format = '#,##0.00'

    ws1.column_dimensions["A"].width = 14
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 40
    ws1.column_dimensions["D"].width = 14
    ws1.column_dimensions["E"].width = 12

    # --- Sheet 2: February ---
    ws2 = wb.create_sheet("February")

    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    feb_data = [
        ["2025-02-02", "Travel", "Uber rides - client visits", 145.60, "Approved"],
        ["2025-02-05", "Equipment", "Monitor stand (x5)", 274.95, "Approved"],
        ["2025-02-07", "Catering", "Valentine's Day team event", 580.00, "Approved"],
        ["2025-02-10", "Software", "Jira Cloud annual plan", 960.00, "Pending"],
        ["2025-02-12", "Marketing", "Social media ad spend - February", 1800.00, "Approved"],
        ["2025-02-15", "Office Supplies", "Desk organizers and filing folders", 112.40, "Approved"],
        ["2025-02-18", "Training", "React workshop - Dev team", 1500.00, "Approved"],
        ["2025-02-20", "Utilities", "Phone service - February billing", 320.00, "Approved"],
        ["2025-02-23", "Travel", "Conference registration - DevSummit 2025", 899.00, "Approved"],
        ["2025-02-25", "Maintenance", "Office cleaning service", 350.00, "Approved"],
        ["2025-02-27", "Equipment", "Wireless mouse and headset combo (x4)", 439.80, "Pending"],
    ]

    for r, row_data in enumerate(feb_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 4:
                cell.number_format = '#,##0.00'

    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 40
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 12

    # --- Sheet 3: March ---
    ws3 = wb.create_sheet("March")

    for col, h in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    mar_data = [
        ["2025-03-03", "Software", "GitHub Enterprise seats (10 additional)", 1440.00, "Approved"],
        ["2025-03-05", "Office Supplies", "Notebooks and pens - bulk order", 156.80, "Approved"],
        ["2025-03-07", "Travel", "Flights to NYC - Q1 board meeting", 1230.00, "Approved"],
        ["2025-03-10", "Catering", "All-hands meeting lunch", 425.50, "Approved"],
        ["2025-03-12", "Marketing", "Email marketing platform - Mailchimp", 299.00, "Approved"],
        ["2025-03-14", "Equipment", "Standing desk converter (x2)", 598.00, "Pending"],
        ["2025-03-17", "Training", "Leadership workshop - managers", 2200.00, "Approved"],
        ["2025-03-19", "Utilities", "Electricity bill - March", 542.30, "Approved"],
        ["2025-03-22", "Travel", "Car rental - Lisa Park site visit", 267.00, "Approved"],
        ["2025-03-25", "Software", "Adobe Creative Cloud team license", 1680.00, "Approved"],
        ["2025-03-27", "Maintenance", "Fire extinguisher inspection", 180.00, "Approved"],
        ["2025-03-30", "Office Supplies", "Toner cartridges - color printer", 289.99, "Approved"],
    ]

    for r, row_data in enumerate(mar_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 4:
                cell.number_format = '#,##0.00'

    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 40
    ws3.column_dimensions["D"].width = 14
    ws3.column_dimensions["E"].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
