"""
Initial Setup: Win/Loss Analysis Dashboard
Task ID: calc_sales_053
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_053'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: WinLoss ---
    ws1 = wb.active
    ws1.title = 'WinLoss'

    # Headers
    headers = ['Deal', 'Rep', 'Result', 'Value', 'Loss Reason']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # Data rows
    data = [
        ['D1', 'Alice', 'Won', 120000, ''],
        ['D2', 'Alice', 'Lost', 85000, 'Price'],
        ['D3', 'Alice', 'Won', 95000, ''],
        ['D4', 'Bob', 'Lost', 150000, 'Competitor'],
        ['D5', 'Bob', 'Won', 200000, ''],
        ['D6', 'Bob', 'Lost', 75000, 'Price'],
        ['D7', 'Alice', 'Won', 60000, ''],
        ['D8', 'Bob', 'Lost', 110000, 'No Budget'],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # --- Sheet 2: Analysis (structure only, NO formulas) ---
    ws2 = wb.create_sheet('Analysis')

    # Rep analysis headers
    rep_headers = ['Rep', 'Deals', 'Wins', 'Win Rate', 'Avg Won Deal', 'Avg Lost Deal', 'Revenue Won']
    for col, h in enumerate(rep_headers, 1):
        ws2.cell(row=1, column=col, value=h)

    # Rep labels only
    ws2.cell(row=2, column=1, value='Alice')
    ws2.cell(row=3, column=1, value='Bob')

    # Loss reason summary headers (columns I-J)
    ws2.cell(row=1, column=9, value='Loss Reason')
    ws2.cell(row=1, column=10, value='Count')

    # Loss reason labels only
    ws2.cell(row=2, column=9, value='Price')
    ws2.cell(row=3, column=9, value='Competitor')
    ws2.cell(row=4, column=9, value='No Budget')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
