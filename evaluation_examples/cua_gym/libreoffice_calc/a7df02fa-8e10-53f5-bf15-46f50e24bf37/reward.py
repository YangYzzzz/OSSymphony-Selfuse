"""
Reward Script: Win/Loss Analysis Dashboard
Task ID: calc_sales_053
Domain: libreoffice_calc
Scoring: Verify formulas in Analysis sheet — rep stats (B2:G3) and loss reason counts (J2:J4)

Components:
  1. Alice deal formulas B2:G2 (0.30 pts) — correct COUNTIF/COUNTIFS/AVERAGEIFS/SUMIFS formulas
  2. Bob deal formulas B3:G3 (0.30 pts) — correct formulas for Bob
  3. Loss reason count formulas J2:J4 (0.20 pts) — COUNTIF formulas for each reason
  4. Win rate formula correctness (0.20 pts) — D2/D3 use ratio of wins/deals
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_053'


def persist_app_state(domain):
    """Save any unsaved LibreOffice state."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def is_formula(val):
    """Check if cell value is a formula string."""
    return isinstance(val, str) and val.startswith('=')


def normalize_formula(f):
    """Normalize formula for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def formula_contains(formula, *keywords):
    """Check if a formula contains all the given keywords (case-insensitive)."""
    norm = normalize_formula(formula)
    return all(kw.upper() in norm for kw in keywords)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Analysis' not in wb.sheetnames:
        print("CRITICAL: 'Analysis' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Analysis']

    # Component 1: Alice stats formulas B2:G2 (0.30 points)
    # B2: COUNTIF for Alice deals, C2: COUNTIFS for Alice wins,
    # D2: Win rate formula, E2: AVERAGEIFS won, F2: AVERAGEIFS lost, G2: SUMIFS won
    try:
        alice_score = 0.0

        # B2: Should be COUNTIF counting Alice in WinLoss rep column
        b2 = ws['B2'].value
        if is_formula(b2) and formula_contains(b2, 'COUNTIF', 'Alice'):
            alice_score += 0.05
            print(f"  PASS: B2 has COUNTIF for Alice: {b2}")
        else:
            print(f"  FAIL: B2 expected COUNTIF for Alice, found: {b2}")

        # C2: Should be COUNTIFS counting Alice + Won
        c2 = ws['C2'].value
        if is_formula(c2) and formula_contains(c2, 'COUNTIF', 'Alice', 'Won'):
            alice_score += 0.05
            print(f"  PASS: C2 has COUNTIFS for Alice+Won: {c2}")
        else:
            print(f"  FAIL: C2 expected COUNTIFS for Alice+Won, found: {c2}")

        # D2: Win rate — should reference C2/B2 or equivalent ratio
        d2 = ws['D2'].value
        if is_formula(d2) and (formula_contains(d2, 'C2', 'B2') or formula_contains(d2, 'COUNTIF')):
            alice_score += 0.05
            print(f"  PASS: D2 has win rate formula: {d2}")
        else:
            print(f"  FAIL: D2 expected win rate formula (C2/B2), found: {d2}")

        # E2: AVERAGEIFS for Alice Won deals
        e2 = ws['E2'].value
        if is_formula(e2) and formula_contains(e2, 'AVERAGEIF', 'Alice', 'Won'):
            alice_score += 0.05
            print(f"  PASS: E2 has AVERAGEIFS for Alice Won: {e2}")
        else:
            print(f"  FAIL: E2 expected AVERAGEIFS for Alice+Won, found: {e2}")

        # F2: AVERAGEIFS for Alice Lost deals
        f2 = ws['F2'].value
        if is_formula(f2) and formula_contains(f2, 'AVERAGEIF', 'Alice', 'Lost'):
            alice_score += 0.05
            print(f"  PASS: F2 has AVERAGEIFS for Alice Lost: {f2}")
        else:
            print(f"  FAIL: F2 expected AVERAGEIFS for Alice+Lost, found: {f2}")

        # G2: SUMIFS for Alice Won revenue
        g2 = ws['G2'].value
        if is_formula(g2) and formula_contains(g2, 'SUMIF', 'Alice', 'Won'):
            alice_score += 0.05
            print(f"  PASS: G2 has SUMIFS for Alice Won: {g2}")
        else:
            print(f"  FAIL: G2 expected SUMIFS for Alice+Won, found: {g2}")

        if alice_score > 0:
            total_score += alice_score
            print(f"PASS: Component 1 — Alice stats ({alice_score:.2f}/0.30 pts)")
        else:
            print(f"FAIL: Component 1 — Alice stats (0.00/0.30 pts)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Bob stats formulas B3:G3 (0.30 points)
    try:
        bob_score = 0.0

        # B3: COUNTIF for Bob deals
        b3 = ws['B3'].value
        if is_formula(b3) and formula_contains(b3, 'COUNTIF', 'Bob'):
            bob_score += 0.05
            print(f"  PASS: B3 has COUNTIF for Bob: {b3}")
        else:
            print(f"  FAIL: B3 expected COUNTIF for Bob, found: {b3}")

        # C3: COUNTIFS for Bob + Won
        c3 = ws['C3'].value
        if is_formula(c3) and formula_contains(c3, 'COUNTIF', 'Bob', 'Won'):
            bob_score += 0.05
            print(f"  PASS: C3 has COUNTIFS for Bob+Won: {c3}")
        else:
            print(f"  FAIL: C3 expected COUNTIFS for Bob+Won, found: {c3}")

        # D3: Win rate formula
        d3 = ws['D3'].value
        if is_formula(d3) and (formula_contains(d3, 'C3', 'B3') or formula_contains(d3, 'COUNTIF')):
            bob_score += 0.05
            print(f"  PASS: D3 has win rate formula: {d3}")
        else:
            print(f"  FAIL: D3 expected win rate formula (C3/B3), found: {d3}")

        # E3: AVERAGEIFS for Bob Won
        e3 = ws['E3'].value
        if is_formula(e3) and formula_contains(e3, 'AVERAGEIF', 'Bob', 'Won'):
            bob_score += 0.05
            print(f"  PASS: E3 has AVERAGEIFS for Bob Won: {e3}")
        else:
            print(f"  FAIL: E3 expected AVERAGEIFS for Bob+Won, found: {e3}")

        # F3: AVERAGEIFS for Bob Lost
        f3 = ws['F3'].value
        if is_formula(f3) and formula_contains(f3, 'AVERAGEIF', 'Bob', 'Lost'):
            bob_score += 0.05
            print(f"  PASS: F3 has AVERAGEIFS for Bob Lost: {f3}")
        else:
            print(f"  FAIL: F3 expected AVERAGEIFS for Bob+Lost, found: {f3}")

        # G3: SUMIFS for Bob Won revenue
        g3 = ws['G3'].value
        if is_formula(g3) and formula_contains(g3, 'SUMIF', 'Bob', 'Won'):
            bob_score += 0.05
            print(f"  PASS: G3 has SUMIFS for Bob Won: {g3}")
        else:
            print(f"  FAIL: G3 expected SUMIFS for Bob+Won, found: {g3}")

        if bob_score > 0:
            total_score += bob_score
            print(f"PASS: Component 2 — Bob stats ({bob_score:.2f}/0.30 pts)")
        else:
            print(f"FAIL: Component 2 — Bob stats (0.00/0.30 pts)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Loss reason count formulas J2:J4 (0.20 points)
    try:
        loss_score = 0.0

        # J2: COUNTIF for "Price"
        j2 = ws['J2'].value
        if is_formula(j2) and formula_contains(j2, 'COUNTIF', 'Price'):
            loss_score += round(0.20 / 3, 4)
            print(f"  PASS: J2 has COUNTIF for Price: {j2}")
        else:
            print(f"  FAIL: J2 expected COUNTIF for Price, found: {j2}")

        # J3: COUNTIF for "Competitor"
        j3 = ws['J3'].value
        if is_formula(j3) and formula_contains(j3, 'COUNTIF', 'Competitor'):
            loss_score += round(0.20 / 3, 4)
            print(f"  PASS: J3 has COUNTIF for Competitor: {j3}")
        else:
            print(f"  FAIL: J3 expected COUNTIF for Competitor, found: {j3}")

        # J4: COUNTIF for "No Budget"
        j4 = ws['J4'].value
        if is_formula(j4) and formula_contains(j4, 'COUNTIF'):
            # Check for "No Budget" or "NoBudget" — might be quoted differently
            j4_norm = normalize_formula(j4)
            if 'NOBUDGET' in j4_norm or 'NO BUDGET' in ws['J4'].value.upper():
                loss_score += round(0.20 / 3, 4)
                print(f"  PASS: J4 has COUNTIF for No Budget: {j4}")
            else:
                print(f"  FAIL: J4 has COUNTIF but missing 'No Budget': {j4}")
        else:
            print(f"  FAIL: J4 expected COUNTIF for No Budget, found: {j4}")

        if loss_score > 0:
            total_score += loss_score
            print(f"PASS: Component 3 — Loss reason counts ({loss_score:.2f}/0.20 pts)")
        else:
            print(f"FAIL: Component 3 — Loss reason counts (0.00/0.20 pts)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Win rate formulas use ratio logic (0.20 points)
    # D2 and D3 should compute wins/total — verify they reference the right cells
    try:
        win_rate_score = 0.0

        d2 = ws['D2'].value
        d3 = ws['D3'].value

        # D2 should be =C2/B2 or equivalent
        if is_formula(d2):
            d2_norm = normalize_formula(d2)
            # Check for C2/B2 pattern or inline COUNTIFS/COUNTIF ratio
            if ('C2/B2' in d2_norm) or ('C2' in d2_norm and 'B2' in d2_norm) or ('COUNTIF' in d2_norm):
                win_rate_score += 0.10
                print(f"  PASS: D2 win rate formula references correct cells: {d2}")
            else:
                print(f"  FAIL: D2 formula doesn't reference C2/B2: {d2}")
        else:
            print(f"  FAIL: D2 is not a formula: {d2}")

        # D3 should be =C3/B3 or equivalent
        if is_formula(d3):
            d3_norm = normalize_formula(d3)
            if ('C3/B3' in d3_norm) or ('C3' in d3_norm and 'B3' in d3_norm) or ('COUNTIF' in d3_norm):
                win_rate_score += 0.10
                print(f"  PASS: D3 win rate formula references correct cells: {d3}")
            else:
                print(f"  FAIL: D3 formula doesn't reference C3/B3: {d3}")
        else:
            print(f"  FAIL: D3 is not a formula: {d3}")

        if win_rate_score > 0:
            total_score += win_rate_score
            print(f"PASS: Component 4 — Win rate formulas ({win_rate_score:.2f}/0.20 pts)")
        else:
            print(f"FAIL: Component 4 — Win rate formulas (0.00/0.20 pts)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
