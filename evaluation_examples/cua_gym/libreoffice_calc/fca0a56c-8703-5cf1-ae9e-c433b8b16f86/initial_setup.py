"""
Initial Setup: Count distinct customer IDs using array formula
Task ID: calc_fmb_array_unique_count_078
Domain: libreoffice_calc
"""

import openpyxl
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_array_unique_count_078'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Orders ---
    ws = wb.active
    ws.title = 'Orders'

    # Headers in row 1
    ws['A1'] = 'Customer ID'
    ws['B1'] = 'Order ID'
    ws['C1'] = 'Amount'

    # Generate 85 distinct customer IDs: C-001 through C-085
    distinct_customers = [f'C-{str(i).zfill(3)}' for i in range(1, 86)]

    # Distribute 300 orders among 85 customers (some get multiple orders)
    # Each customer gets at least 1 order, then distribute remaining 215 randomly
    random.seed(42)
    order_assignments = distinct_customers[:]  # 85 orders (one per customer)
    extra = random.choices(distinct_customers, k=215)  # 215 additional orders
    all_customer_ids = order_assignments + extra  # 300 total
    random.shuffle(all_customer_ids)

    # Write 300 order records in rows 2-301
    for i, cust_id in enumerate(all_customer_ids, 2):
        order_id = f'ORD-{str(i - 1).zfill(5)}'
        amount = round(random.uniform(50.0, 5000.0), 2)
        ws.cell(row=i, column=1, value=cust_id)    # Customer ID
        ws.cell(row=i, column=2, value=order_id)   # Order ID
        ws.cell(row=i, column=3, value=amount)     # Amount

    # Per task context: C2 contains label 'Unique Customers', D2 is empty (target cell)
    # Override C2 with the label text (replaces the amount value placed there)
    ws['C2'] = 'Unique Customers'
    # D2 is left empty — this is where the agent will enter the array formula

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Orders')
    print(f'  Row 1: Headers (Customer ID, Order ID, Amount)')
    print(f'  Rows 2-301: 300 order records with {len(set(all_customer_ids))} distinct customer IDs')
    print(f'  C2: "Unique Customers" (label)')
    print(f'  D2: empty (target cell for array formula)')

create_initial()
