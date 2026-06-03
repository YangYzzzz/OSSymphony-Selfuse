"""
Reward Script: Count distinct customer IDs using array formula in D2
Task ID: calc_fmb_array_unique_count_078
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): D2 contains a formula (not empty/None) — any formula
  Component 2 (0.6): D2 contains a SUMPRODUCT+COUNTIF formula over range A2:A301
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_array_unique_count_078'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Task: Enter an array formula in D2 to count distinct customer IDs in A2:A301.
    Expected formula: =SUMPRODUCT(1/COUNTIF(A2:A301,A2:A301)) evaluating to 85.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook (precondition gate)
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: 'Orders' sheet must exist
    if 'Orders' not in wb.sheetnames:
        print("CRITICAL: 'Orders' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Orders']

    # Component 1: D2 contains a formula (was empty/None in initial state) (0.4 points)
    # This checks the core change: D2 goes from empty to having a formula.
    try:
        d2_value = ws.cell(row=2, column=4).value
        if d2_value is not None and isinstance(d2_value, str) and d2_value.strip().startswith('='):
            print(f"PASS: Component 1 — D2 contains a formula: {repr(d2_value)} (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — D2 expected a formula, found: {repr(d2_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: The formula in D2 uses SUMPRODUCT with COUNTIF over A2:A301 (0.6 points)
    # This verifies the formula is specifically the correct array formula for counting unique values.
    # A valid formula must contain SUMPRODUCT and COUNTIF referencing the correct range A2:A301.
    try:
        d2_value = ws.cell(row=2, column=4).value
        if d2_value is not None and isinstance(d2_value, str):
            formula_upper = d2_value.upper().replace(' ', '')
            has_sumproduct = 'SUMPRODUCT' in formula_upper
            has_countif = 'COUNTIF' in formula_upper
            has_range = 'A2:A301' in formula_upper
            if has_sumproduct and has_countif and has_range:
                print(f"PASS: Component 2 — D2 has SUMPRODUCT+COUNTIF formula over A2:A301: {repr(d2_value)} (0.6 pts)")
                total_score += 0.6
            else:
                missing = []
                if not has_sumproduct:
                    missing.append('SUMPRODUCT')
                if not has_countif:
                    missing.append('COUNTIF')
                if not has_range:
                    missing.append('range A2:A301')
                print(f"FAIL: Component 2 — Missing elements in D2 formula: {missing}, found: {repr(d2_value)}")
        else:
            print(f"FAIL: Component 2 — D2 has no formula to check: {repr(d2_value)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
