"""
Initial Setup: Fix pivot table source range after row deletion
Task ID: calc_pivot_074
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_074'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()

    # --- Sheet 1: CleanedData ---
    ws1 = wb.active
    ws1.title = 'CleanedData'

    # Headers
    headers = ['ID', 'Category', 'Product', 'Sales']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.font = Font(bold=True, size=11, color="FFFFFF")

    # Categories and products for realistic data
    categories = ['Electronics', 'Furniture', 'Office Supplies', 'Clothing', 'Food & Beverage']
    products = {
        'Electronics': ['Laptop Pro X1', 'Wireless Mouse', 'USB-C Hub', '27" Monitor', 'Bluetooth Speaker', 'Webcam HD'],
        'Furniture': ['Standing Desk', 'Ergonomic Chair', 'Bookshelf Oak', 'Filing Cabinet', 'Desk Lamp LED'],
        'Office Supplies': ['Printer Paper A4', 'Sticky Notes Pack', 'Ballpoint Pens 12pk', 'Binder Clips Box', 'Whiteboard Markers'],
        'Clothing': ['Polo Shirt Navy', 'Safety Vest Hi-Vis', 'Work Boots Steel-Toe', 'Lab Coat White', 'Gloves Nitrile'],
        'Food & Beverage': ['Coffee Beans 1kg', 'Green Tea Box', 'Sparkling Water 24pk', 'Protein Bars 12pk', 'Trail Mix Bag'],
    }

    # Generate 85 rows of data (rows 2-86) that sum to 127500
    data_rows = []
    target_total = 127500
    running_total = 0

    for i in range(1, 86):
        cat = categories[i % len(categories)]
        prod_list = products[cat]
        prod = prod_list[i % len(prod_list)]

        if i < 85:
            # Generate sales values that will sum to target
            sales = random.randint(800, 2200)
            running_total += sales
        else:
            # Last row: make it hit the target exactly
            sales = target_total - running_total

        data_rows.append([i, cat, prod, sales])

    for r, row_data in enumerate(data_rows, 2):
        ws1.cell(row=r, column=1, value=row_data[0])  # ID
        ws1.cell(row=r, column=2, value=row_data[1])  # Category
        ws1.cell(row=r, column=3, value=row_data[2])  # Product
        ws1.cell(row=r, column=4, value=row_data[3])  # Sales

    # Set column widths
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 25
    ws1.column_dimensions['D'].width = 12

    # --- Sheet 2: PivotOut ---
    ws2 = wb.create_sheet('PivotOut')

    # Title
    ws2['A1'] = 'Sales Pivot Summary'
    ws2['A1'].font = Font(bold=True, size=14)
    ws2.merge_cells('A1:C1')

    # Source range note (the broken one pointing to deleted rows)
    ws2['A2'] = 'Source: CleanedData!A1:D121'
    ws2['A2'].font = Font(italic=True, color="999999", size=10)

    # Pivot-like headers
    pivot_headers = ['Category', 'Total Sales', 'Count']
    for col, h in enumerate(pivot_headers, 1):
        cell = ws2.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Pivot data rows with #REF! errors to simulate broken references
    for r, cat in enumerate(categories, 5):
        ws2.cell(row=r, column=1, value=cat)
        # Show #REF! errors because original SUMIFS referenced rows 87-121 that no longer exist
        ws2.cell(row=r, column=2, value='=SUMIFS(CleanedData!D2:D121,CleanedData!B2:B121,"' + cat + '")')
        ws2.cell(row=r, column=3, value='=COUNTIFS(CleanedData!B2:B121,"' + cat + '")')

    # Grand total row
    grand_row = 5 + len(categories)
    ws2.cell(row=grand_row, column=1, value='Grand Total')
    ws2.cell(row=grand_row, column=1).font = Font(bold=True)
    ws2.cell(row=grand_row, column=2, value=f'=SUM(B5:B{grand_row - 1})')
    ws2.cell(row=grand_row, column=2).font = Font(bold=True)
    ws2.cell(row=grand_row, column=3, value=f'=SUM(C5:C{grand_row - 1})')
    ws2.cell(row=grand_row, column=3).font = Font(bold=True)

    # Add a thin border around the pivot table area
    thin = Side(style="thin", color="000000")
    for r in range(4, grand_row + 1):
        for c in range(1, 4):
            ws2.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
