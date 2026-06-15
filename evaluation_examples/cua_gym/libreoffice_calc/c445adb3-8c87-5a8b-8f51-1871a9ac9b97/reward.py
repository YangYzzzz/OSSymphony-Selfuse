"""
Reward Script: Fix pivot table source range after row deletion
Task ID: calc_pivot_074
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): Source reference label in A2 updated to D86
  Component 2 (0.4): SUMIFS formulas in B5:B9 reference correct range (D86/B86)
  Component 3 (0.3): COUNTIFS formulas in C5:C9 reference correct range (D86/B86)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_074'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    The task requires fixing pivot table source ranges from D121 to D86
    (because rows 87-121 were deleted from CleanedData).
    All SUMIFS/COUNTIFS in PivotOut must reference row 86 instead of 121.
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: PivotOut sheet must exist
    if 'PivotOut' not in wb.sheetnames:
        print("CRITICAL: 'PivotOut' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['PivotOut']

    # Component 1: Source reference label in A2 updated (0.3 points)
    # Initial has "Source: CleanedData!A1:D121", golden should have "Source: CleanedData!A1:D86"
    # The key change: reference should end at row 86, NOT 121
    try:
        a2_val = ws['A2'].value
        if a2_val and isinstance(a2_val, str):
            # Check that D86 is referenced (correct range) and D121 is NOT
            has_correct = 'D86' in a2_val
            has_old = 'D121' in a2_val
            if has_correct and not has_old:
                print(f"PASS: Component 1 -- A2 source reference updated to D86 (value: {a2_val}) (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 1 -- A2 still references old range or incorrect range (value: {a2_val})")
        else:
            print(f"FAIL: Component 1 -- A2 is empty or not a string (value: {a2_val})")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: SUMIFS formulas in B5:B9 reference correct range (0.4 points)
    # Each of the 5 category SUMIFS should reference D2:D86 and B2:B86 (not D121/B121)
    # Partial credit: 0.08 per correct formula
    try:
        sumifs_correct = 0
        sumifs_total = 5
        categories = ['Electronics', 'Furniture', 'Office Supplies', 'Clothing', 'Food & Beverage']

        for row_idx in range(5, 10):
            cell = ws.cell(row=row_idx, column=2)
            formula = cell.value
            if formula and isinstance(formula, str) and formula.startswith('='):
                # Check that formula references row 86, not 121
                # Valid: references like D2:D86, B2:B86
                has_86 = bool(re.search(r'D\d*:D86', formula, re.IGNORECASE)) or \
                         bool(re.search(r'B\d*:B86', formula, re.IGNORECASE))
                has_121 = bool(re.search(r'[DB]\d*:[DB]121', formula, re.IGNORECASE))

                if has_86 and not has_121:
                    sumifs_correct += 1
                    print(f"  PASS: B{row_idx} SUMIFS correctly references row 86")
                else:
                    print(f"  FAIL: B{row_idx} formula still uses old range or wrong range: {formula}")
            else:
                print(f"  FAIL: B{row_idx} does not contain a formula (value: {formula})")

        comp2_score = 0.4 * (sumifs_correct / sumifs_total)
        if comp2_score > 0:
            print(f"PASS: Component 2 -- {sumifs_correct}/{sumifs_total} SUMIFS formulas correct ({comp2_score:.2f} pts)")
            total_score += comp2_score
        else:
            print(f"FAIL: Component 2 -- No SUMIFS formulas correctly updated")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: COUNTIFS formulas in C5:C9 reference correct range (0.3 points)
    # Each of the 5 category COUNTIFS should reference B2:B86 (not B121)
    # Partial credit: 0.06 per correct formula
    try:
        countifs_correct = 0
        countifs_total = 5

        for row_idx in range(5, 10):
            cell = ws.cell(row=row_idx, column=3)
            formula = cell.value
            if formula and isinstance(formula, str) and formula.startswith('='):
                # Check that formula references row 86, not 121
                has_86 = bool(re.search(r'B\d*:B86', formula, re.IGNORECASE))
                has_121 = bool(re.search(r'B\d*:B121', formula, re.IGNORECASE))

                if has_86 and not has_121:
                    countifs_correct += 1
                    print(f"  PASS: C{row_idx} COUNTIFS correctly references row 86")
                else:
                    print(f"  FAIL: C{row_idx} formula still uses old range or wrong range: {formula}")
            else:
                print(f"  FAIL: C{row_idx} does not contain a formula (value: {formula})")

        comp3_score = 0.3 * (countifs_correct / countifs_total)
        if comp3_score > 0:
            print(f"PASS: Component 3 -- {countifs_correct}/{countifs_total} COUNTIFS formulas correct ({comp3_score:.2f} pts)")
            total_score += comp3_score
        else:
            print(f"FAIL: Component 3 -- No COUNTIFS formulas correctly updated")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
