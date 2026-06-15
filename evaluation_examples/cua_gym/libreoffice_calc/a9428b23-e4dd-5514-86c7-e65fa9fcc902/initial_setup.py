"""
Initial Setup: Spreadsheet with volatile NOW() functions in B1:B1000
Task ID: calc_tbl_040
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_040'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

# Realistic employee/event data for column A
DEPARTMENTS = [
    'Engineering', 'Marketing', 'Sales', 'Finance', 'HR',
    'Operations', 'Legal', 'Support', 'Product', 'Design'
]

FIRST_NAMES = [
    'Sarah', 'Marcus', 'Elena', 'James', 'Priya', 'David', 'Mei',
    'Carlos', 'Fatima', 'Oliver', 'Aisha', 'Thomas', 'Yuki', 'Andre',
    'Sophia', 'Raj', 'Emma', 'Luis', 'Nina', 'Kevin', 'Zara', 'Chen',
    'Amara', 'Viktor', 'Isabel', 'Mohammed', 'Grace', 'Dmitri', 'Leila',
    'Patrick', 'Hana', 'Jorge', 'Chloe', 'Kwame', 'Rosa', 'Ivan',
    'Lena', 'Tariq', 'Mia', 'Sven'
]

LAST_NAMES = [
    'Chen', 'Johnson', 'Petrov', 'Garcia', 'Nakamura', 'Williams',
    'Kumar', 'Mueller', 'Santos', 'O\'Brien', 'Hassan', 'Kim',
    'Johansson', 'Da Silva', 'Wright', 'Okafor', 'Fernandez',
    'Lindqvist', 'Patel', 'Dubois', 'Tanaka', 'Morales', 'Berg',
    'Adeyemi', 'Fischer', 'Lopez', 'Novak', 'Singh', 'Ito', 'Reyes'
]

ACTIONS = [
    'Clock In', 'Clock Out', 'Break Start', 'Break End',
    'Meeting Start', 'Meeting End', 'Shift Start', 'Shift End',
    'Login', 'Logout', 'Task Assigned', 'Task Completed',
    'Report Filed', 'Review Submitted', 'Approval Given'
]

random.seed(42)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: TimeLog ---
    ws = wb.active
    ws.title = 'TimeLog'

    # Header row with formatting
    headers = ['Event Description', 'Timestamp']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows: A = event description, B = =NOW() volatile formula
    for r in range(2, 1002):  # rows 2-1001 = B1 header + B2:B1001 won't work
        # We need B1:B1000 to have =NOW()
        # Since row 1 is header, let's put formulas in B1:B1000
        # Actually task says B1:B1000 - that includes the header row
        pass

    # Re-think: task says B1:B1000 contain =NOW()
    # So row 1 col B is also =NOW(), not a header
    # Let's use row 1 as data, no header row - or use A as header-less

    # Clear what we did
    wb.remove(ws)
    ws = wb.create_sheet('TimeLog', 0)

    # Column A: event descriptions (rows 1-1000)
    # Column B: =NOW() in every cell B1:B1000
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 22

    for r in range(1, 1001):
        first = random.choice(FIRST_NAMES)
        last = random.choice(LAST_NAMES)
        dept = random.choice(DEPARTMENTS)
        action = random.choice(ACTIONS)
        desc = f'{first} {last} - {dept} - {action}'
        ws.cell(row=r, column=1, value=desc)
        ws.cell(row=r, column=2, value='=NOW()')

    # Format column B as datetime
    for r in range(1, 1001):
        ws.cell(row=r, column=2).number_format = 'yyyy-mm-dd hh:mm:ss'

    # --- Sheet 2: Summary ---
    ws2 = wb.create_sheet('Summary')
    ws2['A1'] = 'Total Records'
    ws2['B1'] = 1000
    ws2['A2'] = 'Data Source'
    ws2['B2'] = 'HR Time Tracking System'
    ws2['A3'] = 'Last Updated'
    ws2['B3'] = '=NOW()'
    ws2['B3'].number_format = 'yyyy-mm-dd hh:mm:ss'
    ws2['A4'] = 'Department Count'
    ws2['B4'] = len(DEPARTMENTS)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env['DISPLAY'] = ':0'
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


create_initial()
launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')
