"""
Reward Script: Replace volatile NOW() formulas with static date/time values
Task ID: calc_tbl_040
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): No volatile formulas remain in B1:B1000
  Component 2 (0.3): All B1:B1000 cells contain datetime values
  Component 3 (0.3): All 1000 cells in B1:B1000 have non-None values (data preserved)
"""

import os
import datetime

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_040'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Component 1: No volatile formulas in B1:B1000 (0.4 points)
    # In initial state, all 1000 cells are =NOW(). Task requires removing all formulas.
    try:
        formula_count = 0
        volatile_formulas = []
        for r in range(1, 1001):
            v = ws.cell(row=r, column=2).value
            if isinstance(v, str) and v.startswith('='):
                formula_count += 1
                if len(volatile_formulas) < 5:
                    volatile_formulas.append(f"B{r}: {v}")
        if formula_count == 0:
            print(f"PASS: Component 1 - No formulas found in B1:B1000 (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 - Found {formula_count} formulas in B1:B1000. Examples: {volatile_formulas}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: All B1:B1000 contain datetime values (0.3 points)
    # In initial state, cells contain formula strings, not datetime objects.
    # In golden state, cells contain datetime.datetime objects.
    try:
        datetime_count = 0
        non_datetime_examples = []
        for r in range(1, 1001):
            v = ws.cell(row=r, column=2).value
            if isinstance(v, datetime.datetime):
                datetime_count += 1
            elif len(non_datetime_examples) < 5:
                non_datetime_examples.append(f"B{r}: {repr(v)} (type={type(v).__name__})")

        if datetime_count == 1000:
            print(f"PASS: Component 2 - All 1000 cells contain datetime values (0.3 pts)")
            total_score += 0.3
        elif datetime_count >= 900:
            partial = 0.3 * (datetime_count / 1000)
            if partial > 0:
                print(f"PARTIAL: Component 2 - {datetime_count}/1000 cells are datetime ({partial:.2f} pts). Non-datetime: {non_datetime_examples}")
                total_score += partial
        else:
            print(f"FAIL: Component 2 - Only {datetime_count}/1000 cells are datetime. Examples: {non_datetime_examples}")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: All 1000 cells in B1:B1000 have non-None values (0.3 points)
    # This checks data preservation - the values should not have been deleted.
    # In initial state, cells have formula strings (not None), so this would pass on initial too.
    # BUT: we combine with "not a formula" to make it task-change-specific.
    try:
        valid_count = 0  # non-None AND non-formula
        problem_examples = []
        for r in range(1, 1001):
            v = ws.cell(row=r, column=2).value
            if v is not None and not (isinstance(v, str) and v.startswith('=')):
                valid_count += 1
            elif len(problem_examples) < 5:
                problem_examples.append(f"B{r}: {repr(v)}")

        if valid_count == 1000:
            print(f"PASS: Component 3 - All 1000 cells have static non-None values (0.3 pts)")
            total_score += 0.3
        elif valid_count >= 900:
            partial = 0.3 * (valid_count / 1000)
            if partial > 0:
                print(f"PARTIAL: Component 3 - {valid_count}/1000 cells are valid ({partial:.2f} pts). Problems: {problem_examples}")
                total_score += partial
        else:
            print(f"FAIL: Component 3 - Only {valid_count}/1000 cells have static non-None values. Problems: {problem_examples}")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    final_score = round(min(total_score, 1.0), 1)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist app state before verification
persist_app_state("libreoffice_calc")

# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
