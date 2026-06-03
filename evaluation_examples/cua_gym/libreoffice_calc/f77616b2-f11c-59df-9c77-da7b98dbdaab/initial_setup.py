"""
Initial Setup: Track warehouse labor productivity
Task ID: calc_ops_warehouse_labor_productivity_063
Domain: libreoffice_calc

Creates a spreadsheet with 120 warehouse shift records.
Columns A-G and J are filled; H, I, K, L are left empty (to be computed by the agent).
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date, timedelta
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_warehouse_labor_productivity_063'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'LaborProductivity'

    # --- Headers (Row 1) ---
    headers = [
        'Date',              # A
        'Shift',             # B
        'Operator ID',       # C
        'Hours Worked',      # D
        'Units Picked',      # E
        'Units Packed',      # F
        'Units Shipped',     # G
        'Total Units',       # H — empty (to be calculated)
        'Productivity Units/Hr',  # I — empty (to be calculated)
        'Target Units/Hr',   # J
        'Performance %',     # K — empty (to be calculated)
        'Below Target',      # L — empty (to be calculated)
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.row_dimensions[1].height = 30

    # Column widths
    col_widths = [12, 12, 14, 13, 13, 13, 14, 12, 20, 14, 13, 14]
    col_letters = ['A','B','C','D','E','F','G','H','I','J','K','L']
    for letter, width in zip(col_letters, col_widths):
        ws.column_dimensions[letter].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # --- Operators ---
    operators = [
        'WH-001', 'WH-002', 'WH-003', 'WH-004', 'WH-005',
        'WH-006', 'WH-007', 'WH-008', 'WH-009', 'WH-010',
    ]

    # Shift definitions: name -> (target_units_per_hr, hours, pick_range, pack_range, ship_range)
    shifts = {
        'Morning':   {'target': 85, 'hours': 8.0,
                      'pick': (200, 320), 'pack': (180, 310), 'ship': (160, 280)},
        'Afternoon': {'target': 80, 'hours': 8.0,
                      'pick': (190, 300), 'pack': (170, 290), 'ship': (150, 270)},
        'Night':     {'target': 70, 'hours': 7.0,
                      'pick': (160, 260), 'pack': (150, 240), 'ship': (130, 220)},
    }
    shift_names = ['Morning', 'Afternoon', 'Night']

    # Generate 120 records spread over ~40 dates
    start_date = date(2025, 1, 6)  # Monday

    rows = []
    for i in range(120):
        day_offset = i // 3        # 3 shifts per day
        shift_idx = i % 3
        record_date = start_date + timedelta(days=day_offset)
        # Skip weekends (Sat=5, Sun=6)
        weekday = record_date.weekday()
        if weekday >= 5:
            record_date += timedelta(days=2 - (weekday - 5))

        shift_name = shift_names[shift_idx]
        shift_info = shifts[shift_name]
        operator = operators[i % len(operators)]

        hours = shift_info['hours']
        # Occasionally vary hours slightly
        hours_var = round(hours + random.choice([-0.5, 0, 0, 0.5]), 1)

        picked = random.randint(*shift_info['pick'])
        packed = random.randint(*shift_info['pack'])
        shipped = random.randint(*shift_info['ship'])

        target = shift_info['target']

        rows.append([
            record_date,   # A - Date
            shift_name,    # B - Shift
            operator,      # C - Operator ID
            hours_var,     # D - Hours Worked
            picked,        # E - Units Picked
            packed,        # F - Units Packed
            shipped,       # G - Units Shipped
            None,          # H - Total Units (empty)
            None,          # I - Productivity Units/Hr (empty)
            target,        # J - Target Units/Hr
            None,          # K - Performance % (empty)
            None,          # L - Below Target (empty)
        ])

    # Write data rows
    for r_idx, row_data in enumerate(rows, 2):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if c_idx == 1 and val is not None:
                cell.number_format = 'yyyy-mm-dd'
            elif c_idx == 4:
                cell.number_format = '0.0'
            elif c_idx in (10,):  # Target
                cell.number_format = '0'

    # Auto-filter on header row
    ws.auto_filter.ref = 'A1:L1'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: LaborProductivity')
    print(f'  Data rows: 120 (rows 2-121)')
    print(f'  Columns A-G and J filled; H, I, K, L empty (to be calculated)')

create_initial()
