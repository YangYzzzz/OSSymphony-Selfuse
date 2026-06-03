"""
Reward Script: Track warehouse labor productivity
Task ID: calc_ops_warehouse_labor_productivity_063
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): H2:H121 formulas = E+F+G (Total Units)
  Component 2 (0.25): I2:I121 formulas = H/D (Productivity Units/Hr)
  Component 3 (0.20): K2:K121 formulas = I/J (Performance %), formatted as percentage
  Component 4 (0.10): L2:L121 formulas = IF(K<1, "BELOW TARGET", "OK")
  Component 5 (0.05): Conditional formatting on L2:L121 (red for BELOW TARGET)
  Component 6 (0.10): Summary section with AVERAGEIF by operator and COUNTIFS per operator/shift
"""

import os

import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_ops_warehouse_labor_productivity_063'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet 'LaborProductivity' must exist
    if 'LaborProductivity' not in wb.sheetnames:
        print("CRITICAL: Sheet 'LaborProductivity' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['LaborProductivity']

    # Component 1: H2:H121 contain formulas summing E+F+G (Total Units) (0.30 points)
    # These should be None in initial file, formulas in golden file
    try:
        h_formula_count = 0
        h_formula_correct = 0
        for row in range(2, 122):  # rows 2-121 = 120 rows
            val = ws.cell(row=row, column=8).value  # column H
            if val is not None:
                h_formula_count += 1
                # Normalize formula — check it sums E, F, G for the same row
                val_str = str(val).upper().replace(" ", "")
                e_col = f"E{row}"
                f_col = f"F{row}"
                g_col = f"G{row}"
                # Accept =E2+F2+G2 style
                if (e_col in val_str and f_col in val_str and g_col in val_str and "+" in val_str):
                    h_formula_correct += 1

        if h_formula_correct == 120:
            print(f"PASS: Component 1 — H2:H121 all have correct Total Units formulas (=En+Fn+Gn) ({h_formula_correct}/120) (0.30 pts)")
            total_score += 0.30
        elif h_formula_correct >= 60:
            # Partial credit for partial completion
            partial = round(0.30 * (h_formula_correct / 120), 2)
            print(f"PARTIAL: Component 1 — H formulas: {h_formula_correct}/120 correct (partial {partial} pts)")
            if partial > 0:
                total_score += partial
        else:
            print(f"FAIL: Component 1 — H2:H121 Total Units formulas: only {h_formula_correct}/120 correct (found {h_formula_count} non-empty)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: I2:I121 contain formulas = H/D (Productivity Units/Hr) (0.25 points)
    try:
        i_formula_correct = 0
        for row in range(2, 122):
            val = ws.cell(row=row, column=9).value  # column I
            if val is not None:
                val_str = str(val).upper().replace(" ", "")
                h_col = f"H{row}"
                d_col = f"D{row}"
                # Accept =H2/D2 style
                if h_col in val_str and d_col in val_str and "/" in val_str:
                    i_formula_correct += 1

        if i_formula_correct == 120:
            print(f"PASS: Component 2 — I2:I121 all have correct Productivity formulas (=Hn/Dn) ({i_formula_correct}/120) (0.25 pts)")
            total_score += 0.25
        elif i_formula_correct >= 60:
            partial = round(0.25 * (i_formula_correct / 120), 2)
            print(f"PARTIAL: Component 2 — I formulas: {i_formula_correct}/120 correct (partial {partial} pts)")
            if partial > 0:
                total_score += partial
        else:
            print(f"FAIL: Component 2 — I2:I121 Productivity formulas: only {i_formula_correct}/120 correct")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: K2:K121 contain formulas = I/J (Performance %) and formatted as percentage (0.20 points)
    try:
        k_formula_correct = 0
        k_format_correct = 0
        for row in range(2, 122):
            cell = ws.cell(row=row, column=11)  # column K
            val = cell.value
            if val is not None:
                val_str = str(val).upper().replace(" ", "")
                i_col = f"I{row}"
                j_col = f"J{row}"
                # Accept =I2/J2 style
                if i_col in val_str and j_col in val_str and "/" in val_str:
                    k_formula_correct += 1
                # Check number format is percentage-like
                nf = cell.number_format or ""
                if "%" in nf:
                    k_format_correct += 1

        if k_formula_correct == 120:
            print(f"PASS: Component 3 — K2:K121 all have correct Performance % formulas (=In/Jn) ({k_formula_correct}/120) (0.15 pts)")
            total_score += 0.15
            if k_format_correct >= 100:
                print(f"PASS: Component 3b — K column percentage number format applied ({k_format_correct}/120) (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 3b — K column percentage format only applied to {k_format_correct}/120 cells")
        elif k_formula_correct >= 60:
            partial = round(0.15 * (k_formula_correct / 120), 2)
            print(f"PARTIAL: Component 3 — K formulas: {k_formula_correct}/120 correct (partial {partial} pts)")
            if partial > 0:
                total_score += partial
        else:
            print(f"FAIL: Component 3 — K2:K121 Performance % formulas: only {k_formula_correct}/120 correct")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: L2:L121 contain IF formulas classifying BELOW TARGET / OK (0.10 points)
    try:
        l_formula_correct = 0
        for row in range(2, 122):
            val = ws.cell(row=row, column=12).value  # column L
            if val is not None:
                val_str = str(val).upper().replace(" ", "")
                # Accept =IF(K2<1,"BELOW TARGET","OK") or similar
                if "IF(" in val_str and "BELOWTARGET" in val_str.replace('"', "").replace("'", "") and "OK" in val_str:
                    l_formula_correct += 1

        if l_formula_correct == 120:
            print(f"PASS: Component 4 — L2:L121 all have BELOW TARGET/OK IF formulas ({l_formula_correct}/120) (0.10 pts)")
            total_score += 0.10
        elif l_formula_correct >= 60:
            partial = round(0.10 * (l_formula_correct / 120), 2)
            print(f"PARTIAL: Component 4 — L formulas: {l_formula_correct}/120 correct (partial {partial} pts)")
            if partial > 0:
                total_score += partial
        else:
            print(f"FAIL: Component 4 — L2:L121 BELOW TARGET formulas: only {l_formula_correct}/120 correct")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Conditional formatting on L2:L121 (red fill for BELOW TARGET) (0.05 points)
    try:
        cf_rules = ws.conditional_formatting._cf_rules

        def has_red_fill_on_L(cf_rules_dict):
            """Return True if any CF rule on L column has a red fill."""
            for cf_range_obj, rules in cf_rules_dict.items():
                if "L" in str(cf_range_obj).upper():
                    for rule in rules:
                        if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                            fill_color = rule.dxf.fill.fgColor.rgb if rule.dxf.fill.fgColor else ""
                            # Accept red fill: FFFF0000
                            if fill_color and "FF0000" in fill_color.upper():
                                return True
            return False

        cf_l_col_found = any("L" in str(cf_range).upper() for cf_range in cf_rules.keys())
        cf_red_confirmed = has_red_fill_on_L(cf_rules)

        if cf_l_col_found and cf_red_confirmed:
            print(f"PASS: Component 5 — Conditional formatting on L column with red fill (0.05 pts)")
            total_score += 0.05
        elif cf_l_col_found:
            print(f"FAIL: Component 5 — Conditional formatting found on L but red fill not confirmed")
        else:
            print(f"FAIL: Component 5 — No conditional formatting found on L column")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Summary section — AVERAGEIF by operator, COUNTIFS below-target (0.10 points)
    # Check: row ~124 has a summary title, rows ~127-136 have AVERAGEIF formulas for operators
    # and rows ~140-142 have AVERAGEIF formulas for shifts
    try:
        # Find summary header: look for 'PRODUCTIVITY SUMMARY' or similar in column A after row 121
        summary_header_row = None
        for row in range(122, 160):
            val = ws.cell(row=row, column=1).value
            if val and "PRODUCTIVITY" in str(val).upper() and "SUMMARY" in str(val).upper():
                summary_header_row = row
                break

        # Find AVERAGEIF and COUNTIFS formulas in column B/C after summary header
        averageif_count = 0
        countifs_count = 0
        if summary_header_row is not None:
            for row in range(summary_header_row, summary_header_row + 30):
                b_val = ws.cell(row=row, column=2).value
                c_val = ws.cell(row=row, column=3).value
                if b_val and "AVERAGEIF" in str(b_val).upper():
                    averageif_count += 1
                if c_val and "COUNTIFS" in str(c_val).upper():
                    countifs_count += 1

        if summary_header_row is not None and averageif_count >= 10 and countifs_count >= 10:
            print(f"PASS: Component 6 — Summary section at row {summary_header_row}: {averageif_count} AVERAGEIF + {countifs_count} COUNTIFS formulas (0.10 pts)")
            total_score += 0.10
        elif summary_header_row is not None and (averageif_count >= 5 or countifs_count >= 5):
            print(f"PARTIAL: Component 6 — Summary section partially complete: {averageif_count} AVERAGEIF, {countifs_count} COUNTIFS (partial 0.05 pts)")
            total_score += 0.05
        elif summary_header_row is not None:
            print(f"FAIL: Component 6 — Summary header found at row {summary_header_row} but AVERAGEIF={averageif_count}, COUNTIFS={countifs_count}")
        else:
            print(f"FAIL: Component 6 — No productivity summary section found after row 121")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
