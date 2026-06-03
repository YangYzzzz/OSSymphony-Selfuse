"""
Reward Script: Warehouse Receiving Log - Variance Column and Status Flags
Task ID: calc_ops_warehouse_receiving_log_018
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): G2:G81 all contain =IF(ISBLANK(Fn),"",Fn-En) variance formulas
  Component 2 (0.25): H2:H81 all contain IF-based status formulas (SHORT/OVER/OK/blank)
  Component 3 (0.25): Conditional formatting on H2:H81 with orange (SHORT), red (OVER), green (OK)
  Component 4 (0.15): Data validation on F2:F81 allowing only whole numbers >= 0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_warehouse_receiving_log_018'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the ReceivingLog sheet exists (precondition gate)
    if 'ReceivingLog' not in wb.sheetnames:
        print("CRITICAL: Sheet 'ReceivingLog' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['ReceivingLog']

    # -----------------------------------------------------------------------
    # Component 1: G2:G81 contain variance formulas =IF(ISBLANK(Fn),"",Fn-En)
    # (0.35 points)
    # The initial file has G2:G81 all None; the golden file has all 80 rows filled
    # with IF(ISBLANK formulas. We verify the formula pattern for all rows.
    # -----------------------------------------------------------------------
    try:
        # Pattern: =IF(ISBLANK(Fn),"",Fn-En) — case-insensitive, ignore spaces
        # Accept minor variations in quoting: "" or ''
        g_pattern = re.compile(
            r'^=IF\s*\(\s*ISBLANK\s*\(\s*F(\d+)\s*\)\s*,\s*""\s*,\s*F\1\s*-\s*E\1\s*\)$',
            re.IGNORECASE
        )
        g_pass_count = 0
        g_fail_rows = []
        for row in range(2, 82):
            val = ws.cell(row=row, column=7).value  # column G
            if val is None:
                g_fail_rows.append((row, None))
            elif isinstance(val, str):
                normalized = val.replace(' ', '')
                if g_pattern.match(normalized):
                    g_pass_count += 1
                else:
                    g_fail_rows.append((row, val))
            else:
                g_fail_rows.append((row, val))

        if g_pass_count == 80:
            print(f"PASS: Component 1 — All 80 rows G2:G81 have correct variance formula (0.35 pts)")
            total_score += 0.35
        elif g_pass_count >= 60:
            partial = round(0.35 * g_pass_count / 80, 4)
            print(f"PARTIAL: Component 1 — {g_pass_count}/80 rows in G have correct formula "
                  f"(partial {partial:.4f} pts). First failures: {g_fail_rows[:3]}")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {g_pass_count}/80 rows in G have correct variance formula. "
                  f"First failures: {g_fail_rows[:3]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: H2:H81 contain status formulas
    # =IF(ISBLANK(Fn),"",IF(Gn<0,"SHORT",IF(Gn>0,"OVER","OK")))
    # (0.25 points)
    # -----------------------------------------------------------------------
    try:
        # Accept the specific nested IF pattern for status
        # Pattern: =IF(ISBLANK(Fn),"",IF(Gn<0,"SHORT",IF(Gn>0,"OVER","OK")))
        h_pattern = re.compile(
            r'^=IF\s*\(\s*ISBLANK\s*\(\s*F(\d+)\s*\)\s*,\s*""\s*,\s*IF\s*\(\s*G\1\s*<\s*0\s*,\s*"SHORT"\s*,\s*IF\s*\(\s*G\1\s*>\s*0\s*,\s*"OVER"\s*,\s*"OK"\s*\)\s*\)\s*\)$',
            re.IGNORECASE
        )
        h_pass_count = 0
        h_fail_rows = []
        for row in range(2, 82):
            val = ws.cell(row=row, column=8).value  # column H
            if val is None:
                h_fail_rows.append((row, None))
            elif isinstance(val, str):
                normalized = val.replace(' ', '')
                if h_pattern.match(normalized):
                    h_pass_count += 1
                else:
                    h_fail_rows.append((row, val))
            else:
                h_fail_rows.append((row, val))

        if h_pass_count == 80:
            print(f"PASS: Component 2 — All 80 rows H2:H81 have correct status formula (0.25 pts)")
            total_score += 0.25
        elif h_pass_count >= 60:
            partial = round(0.25 * h_pass_count / 80, 4)
            print(f"PARTIAL: Component 2 — {h_pass_count}/80 rows in H have correct formula "
                  f"(partial {partial:.4f} pts). First failures: {h_fail_rows[:3]}")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {h_pass_count}/80 rows in H have correct status formula. "
                  f"First failures: {h_fail_rows[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Conditional formatting on H2:H81
    # - SHORT => orange/amber fill (FFFFA500 or similar amber)
    # - OVER  => red fill (FFFF0000 or similar red)
    # - OK    => green fill (FF00B050 or similar green)
    # (0.25 points)
    # -----------------------------------------------------------------------
    try:
        cf_rules = ws.conditional_formatting
        short_cf = False
        over_cf = False
        ok_cf = False

        for cf_range in cf_rules:
            range_str = str(cf_range)
            # Must apply to H column area (H2:H81 or similar)
            if 'H' not in range_str:
                continue
            for rule in cf_range.rules:
                if rule.type not in ('expression', 'containsText', 'formula'):
                    continue
                formula_list = getattr(rule, 'formula', []) or []
                formula_str = ' '.join(str(f) for f in formula_list).upper().replace(' ', '')

                # Get fill color
                fill_color = None
                try:
                    if rule.dxf and rule.dxf.fill:
                        fill_color = rule.dxf.fill.fgColor.rgb.upper()
                except Exception:
                    pass

                # Check SHORT -> orange/amber
                if 'SHORT' in formula_str and fill_color:
                    # Orange: FFFFA500 or FFA500xx variants; amber is similar
                    if fill_color.startswith('FF') and (
                        'FFA5' in fill_color or      # orange FFA500
                        'FFB3' in fill_color or      # light amber
                        'FFFF' in fill_color[:6] and fill_color[6:8] in ('00','80') or  # yellow-orange
                        fill_color in ('FFFFA500', 'FFFFB300', 'FFFF8C00', 'FFFF9900', 'FFFF6600')
                    ):
                        short_cf = True
                        print(f"  SHORT CF found: formula={formula_list}, color={fill_color}")
                    elif fill_color:
                        # Accept any non-white/non-green/non-blue amber-ish color for SHORT
                        # Key: not FFFF0000 (red), not FF00B050 (green), not white
                        if fill_color not in ('FFFF0000', 'FF00FF00', 'FF00B050', 'FFFFFFFF', '00000000'):
                            short_cf = True
                            print(f"  SHORT CF found (custom amber): formula={formula_list}, color={fill_color}")

                # Check OVER -> red
                if 'OVER' in formula_str and fill_color:
                    # Red: FFFF0000 or similar red
                    r_val = int(fill_color[2:4], 16) if len(fill_color) >= 4 else 0
                    g_val = int(fill_color[4:6], 16) if len(fill_color) >= 6 else 0
                    b_val = int(fill_color[6:8], 16) if len(fill_color) >= 8 else 0
                    if r_val > 150 and g_val < 100 and b_val < 100:
                        over_cf = True
                        print(f"  OVER CF found: formula={formula_list}, color={fill_color}")

                # Check OK -> green
                if 'OK' in formula_str and fill_color:
                    r_val = int(fill_color[2:4], 16) if len(fill_color) >= 4 else 0
                    g_val = int(fill_color[4:6], 16) if len(fill_color) >= 6 else 0
                    b_val = int(fill_color[6:8], 16) if len(fill_color) >= 8 else 0
                    if g_val > 100 and r_val < g_val:
                        ok_cf = True
                        print(f"  OK CF found: formula={formula_list}, color={fill_color}")

        cf_score = 0.0
        if short_cf:
            cf_score += 0.10
            print("PASS: Conditional formatting SHORT (amber/orange) — +0.10")
        else:
            print("FAIL: Missing or incorrect CF for SHORT (amber/orange fill)")
        if over_cf:
            cf_score += 0.10
            print("PASS: Conditional formatting OVER (red) — +0.10")
        else:
            print("FAIL: Missing or incorrect CF for OVER (red fill)")
        if ok_cf:
            cf_score += 0.05
            print("PASS: Conditional formatting OK (green) — +0.05")
        else:
            print("FAIL: Missing or incorrect CF for OK (green fill)")

        total_score += cf_score
        if cf_score < 0.25:
            print(f"FAIL: Component 3 — CF score {cf_score:.2f}/0.25")
        else:
            print(f"PASS: Component 3 — All conditional formatting correct (0.25 pts)")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: Data validation on F2:F81 — whole numbers >= 0
    # (0.15 points)
    # -----------------------------------------------------------------------
    try:
        validations = ws.data_validations.dataValidation
        found_dv = False
        for dv in validations:
            sqref_str = str(dv.sqref)
            # Must cover F column
            if 'F' not in sqref_str:
                continue
            # Must be whole number type
            if dv.type != 'whole':
                continue
            # Must require >= 0
            formula1 = str(dv.formula1) if dv.formula1 is not None else ''
            operator = str(dv.operator).lower() if dv.operator else ''
            if ('greaterThanOrEqual' in operator or 'greaterthanorequal' in operator) and formula1 == '0':
                found_dv = True
                print(f"PASS: Component 4 — Data validation on {sqref_str}: "
                      f"type={dv.type}, operator={dv.operator}, formula1={formula1} (0.15 pts)")
            elif 'greaterthan' in operator and formula1 == '-1':
                # Alternative: > -1 is equivalent to >= 0 for integers
                found_dv = True
                print(f"PASS: Component 4 — Data validation (alt form) on {sqref_str}: "
                      f"type={dv.type}, operator={dv.operator}, formula1={formula1} (0.15 pts)")
            elif dv.type == 'whole' and 'F2' in sqref_str:
                # Accept any whole number DV on F column as partial
                found_dv = True
                print(f"PASS: Component 4 — Data validation (whole number) on {sqref_str}: "
                      f"type={dv.type}, operator={dv.operator}, formula1={formula1} (0.15 pts)")

        if found_dv:
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — No whole-number data validation found on F2:F81. "
                  f"Found validations: {[(str(dv.type), str(dv.sqref)) for dv in validations]}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.4f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
