"""
Initial Setup: Warehouse Receiving Log
Task ID: calc_ops_warehouse_receiving_log_018
Domain: libreoffice_calc

Creates the ReceivingLog spreadsheet with:
- Headers in row 1 (columns A-J)
- 80 receiving entries in rows 2-81 (A-E, I filled; F partial; G-H empty)
- No formulas, no conditional formatting, no data validation
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date, timedelta
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_warehouse_receiving_log_018'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ReceivingLog'

    # --- Row 1: Headers ---
    headers = [
        'Received Date',   # A
        'PO Number',       # B
        'SKU',             # C
        'Product',         # D
        'Qty Ordered',     # E
        'Qty Received',    # F  (data entry column)
        'Variance',        # G  (to be filled by task)
        'Status',          # H  (to be filled by task)
        'Receiver Name',   # I
        'Notes',           # J
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # --- Realistic data sets ---
    products = [
        ('SKU-1001', 'Industrial Safety Gloves (Box/100)', 200, 250),
        ('SKU-1002', 'Warehouse Pallet Wrap (Roll)', 50, 150),
        ('SKU-1003', 'Steel Toe Work Boots (Pair)', 20, 60),
        ('SKU-1004', 'Hard Hat Standard Yellow', 30, 80),
        ('SKU-1005', 'High-Vis Vest Size M', 100, 300),
        ('SKU-1006', 'Forklift Battery Acid (5L)', 10, 40),
        ('SKU-1007', 'Stretch Wrap Film 500mm', 80, 200),
        ('SKU-1008', 'Cargo Straps 5T (Set)', 15, 50),
        ('SKU-1009', 'Floor Marking Tape Yellow', 40, 120),
        ('SKU-1010', 'Hand Truck 300kg Capacity', 5, 20),
        ('SKU-1011', 'Hydraulic Pallet Jack 2T', 3, 10),
        ('SKU-1012', 'Safety Earmuffs NRR25', 60, 180),
        ('SKU-1013', 'Chemical Resistant Apron', 25, 75),
        ('SKU-1014', 'Dust Mask FFP2 (Box/20)', 150, 400),
        ('SKU-1015', 'First Aid Kit Type C', 8, 25),
        ('SKU-1016', 'Anti-Fatigue Mat 90x60cm', 12, 35),
        ('SKU-1017', 'Warehouse Label Printer Tape', 30, 90),
        ('SKU-1018', 'Barcode Scanner Handheld', 4, 12),
        ('SKU-1019', 'Loading Bay Dock Seal', 2, 8),
        ('SKU-1020', 'Shrink Wrap Bags Large', 500, 1000),
    ]

    receivers = [
        'James Whitfield', 'Priya Nair', 'Chen Jianfeng', 'Maria Santos',
        'Daniel Okonkwo', 'Rachel Thornton', 'Ahmed Al-Rashid', 'Linh Nguyen',
        'Tobias Becker', 'Fatima Malik',
    ]

    suppliers = [
        'GlobalSafe Supplies', 'PrimeWare Co', 'Industrial Direct', 'SafetyFirst Ltd',
        'Workwear Depot', 'PackagingPro', 'Ergonomic Solutions', 'LogiTech Warehouse',
    ]

    notes_options = [
        'Delivered on time', 'Pallets damaged on arrival', 'Partial delivery accepted',
        'Checked against PO', 'Driver signature obtained', 'Cold chain maintained',
        'No discrepancies noted', 'Outer packaging torn - contents OK',
        '', '', '', '', '',  # Many blank notes
    ]

    # Generate 80 entries
    base_date = date(2025, 10, 1)

    # 60 rows have F filled, 20 rows have F empty (not yet received)
    filled_f_indices = set(random.sample(range(2, 82), 60))

    for i in range(80):
        row = i + 2
        entry_date = base_date + timedelta(days=random.randint(0, 120))
        po_number = f'PO-{2025 if entry_date.year == 2025 else 2026}-{random.randint(10000, 99999)}'
        sku, product, qty_min, qty_max = random.choice(products)
        qty_ordered = random.randint(qty_min, qty_max)
        receiver = random.choice(receivers)
        note = random.choice(notes_options)

        ws.cell(row=row, column=1, value=entry_date)
        ws.cell(row=row, column=1).number_format = 'yyyy-mm-dd'
        ws.cell(row=row, column=2, value=po_number)
        ws.cell(row=row, column=3, value=sku)
        ws.cell(row=row, column=4, value=product)
        ws.cell(row=row, column=5, value=qty_ordered)
        # Column F (Qty Received) - only fill some rows
        if row in filled_f_indices:
            # Some over-shipments, some short, some exact
            variance_type = random.choice(['short', 'short', 'exact', 'exact', 'over'])
            if variance_type == 'short':
                qty_received = max(0, qty_ordered - random.randint(1, max(1, qty_ordered // 5)))
            elif variance_type == 'over':
                qty_received = qty_ordered + random.randint(1, max(1, qty_ordered // 10))
            else:
                qty_received = qty_ordered
            ws.cell(row=row, column=6, value=qty_received)
        # Columns G and H are intentionally left empty
        # Column I: Receiver Name
        ws.cell(row=row, column=9, value=receiver)
        # Column J: Notes
        if note:
            ws.cell(row=row, column=10, value=note)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 38
    ws.column_dimensions['E'].width = 13
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 11
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 30

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: ReceivingLog')
    print(f'  Rows: 80 data rows (rows 2-81)')
    print(f'  Columns filled: A-F (partial), I (receiver), J (notes)')
    print(f'  G and H intentionally empty (no formulas)')

create_initial()
