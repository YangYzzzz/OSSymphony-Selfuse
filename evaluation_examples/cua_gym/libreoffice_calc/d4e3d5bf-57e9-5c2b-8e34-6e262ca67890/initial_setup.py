"""
Initial Setup: Temperature log spreadsheet with January daily data.
Task ID: calc_fmb_min_max_005
Domain: libreoffice_calc

Creates a spreadsheet with one sheet 'Temperature Log' containing 31 days of
January temperature, humidity and weather condition data. E2 and E3 are left
empty as the target cells for MAX/MIN formulas.
"""

import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_min_max_005'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Temperature Log'

    # --- Row 1: Headers ---
    ws['A1'] = 'Date'
    ws['B1'] = 'Temperature (\u00b0C)'
    ws['C1'] = 'Humidity (%)'
    ws['D1'] = 'Condition'

    # Bold headers
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}1'].font = Font(bold=True)

    # --- Rows 2-32: 31 days of January data ---
    # Temperatures exactly as specified in context
    temperatures = [
        3.2, 1.8, -0.5, 2.1, 4.7, 6.2, 5.8, 3.4, 2.9, -1.2,
        -2.8, 0.4, 1.6, 3.8, 5.1, 7.3, 8.9, 7.6, 6.4, 4.2,
        2.7, 1.1, -0.3, 1.9, 3.6, 5.5, 6.8, 8.1, 9.4, 8.7, 7.2
    ]

    # Realistic humidity values (%) for a January in a temperate city
    humidity = [
        72, 68, 75, 70, 65, 63, 67, 71, 74, 80,
        82, 78, 76, 69, 64, 60, 58, 61, 66, 70,
        73, 77, 79, 74, 68, 62, 59, 57, 55, 58, 62
    ]

    # Realistic weather conditions
    conditions = [
        'Cloudy', 'Overcast', 'Snow', 'Partly Cloudy', 'Clear',
        'Clear', 'Cloudy', 'Cloudy', 'Overcast', 'Sleet',
        'Snow', 'Snow', 'Overcast', 'Partly Cloudy', 'Clear',
        'Sunny', 'Sunny', 'Clear', 'Clear', 'Partly Cloudy',
        'Cloudy', 'Overcast', 'Foggy', 'Partly Cloudy', 'Clear',
        'Clear', 'Sunny', 'Sunny', 'Sunny', 'Clear', 'Clear'
    ]

    for day in range(1, 32):
        row = day + 1
        # Date: January 2025
        ws.cell(row=row, column=1, value=f'2025-01-{day:02d}')
        ws.cell(row=row, column=2, value=temperatures[day - 1])
        ws.cell(row=row, column=3, value=humidity[day - 1])
        ws.cell(row=row, column=4, value=conditions[day - 1])

    # --- Labels in D2 and D3 as specified in context ---
    ws['D2'] = 'Max Temp'
    ws['D3'] = 'Min Temp'

    # --- E2 and E3 are intentionally left EMPTY (task target cells) ---

    # Set column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Temperature Log')
    print(f'  Rows: 1 header + 31 data rows (Jan 1-31, 2025)')
    print(f'  E2 and E3: empty (awaiting MAX/MIN formulas)')


create_initial()
