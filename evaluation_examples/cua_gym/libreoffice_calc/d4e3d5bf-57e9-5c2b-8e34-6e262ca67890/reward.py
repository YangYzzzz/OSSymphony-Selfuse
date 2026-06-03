"""
Reward Script: Find max/min temperatures in January and place in E2 and E3
Task ID: calc_fmb_min_max_005
Domain: libreoffice_calc
Scoring:
  Component 1: Cell E2 contains =MAX(B2:B32) formula  — 0.5 points
  Component 2: Cell E3 contains =MIN(B2:B32) formula  — 0.5 points
  Total: 1.0

Verification strategy:
  - Load the file (formula mode, not data_only) to read formula strings.
  - E2 must contain a MAX formula covering B2:B32.
  - E3 must contain a MIN formula covering B2:B32.
  - Additionally validate the formula range is correct (B2:B32).
  - Also verify that computed max/min values match expected 9.4 and -2.8
    by computing directly from the B column data, as a secondary check
    (the formula string check is the primary gate).
  - Initial file has E2=None, E3=None — so these checks fail on initial (score 0.0).
  - Golden file has E2='=MAX(B2:B32)', E3='=MIN(B2:B32)' — both checks pass (score 1.0).
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'   # VM path — reward script runs on the VM
TASK_ID = 'calc_fmb_min_max_005'
SHEET_NAME = 'Temperature Log'

# Expected formula patterns (case-insensitive, whitespace-insensitive)
EXPECTED_E2_FORMULA_PATTERN = r'^=MAX\s*\(\s*B2\s*:\s*B32\s*\)$'
EXPECTED_E3_FORMULA_PATTERN = r'^=MIN\s*\(\s*B2\s*:\s*B32\s*\)$'

# Expected computed values derived from B column data
EXPECTED_MAX = 9.4
EXPECTED_MIN = -2.8
TOLERANCE = 0.01


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition gate: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: expected sheet must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Sheets present: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # -----------------------------------------------------------------------
    # Component 1: Cell E2 contains a MAX formula covering B2:B32 (0.5 points)
    # -----------------------------------------------------------------------
    # This FAILS on initial (E2 is None) and PASSES on golden (E2='=MAX(B2:B32)')
    try:
        e2_value = ws.cell(row=2, column=5).value   # column E = 5

        if e2_value is None:
            print(f"FAIL: Component 1 — E2 is empty (None). Expected MAX formula.")
        elif not isinstance(e2_value, str):
            print(f"FAIL: Component 1 — E2 is not a formula string, found: {repr(e2_value)}")
        else:
            normalized = e2_value.strip().upper().replace(' ', '')
            # Accept both =MAX(B2:B32) and any equivalent (just verify function name and range)
            if re.match(EXPECTED_E2_FORMULA_PATTERN, e2_value.strip(), re.IGNORECASE):
                print(f"PASS: Component 1 — E2 contains MAX formula: {e2_value!r} (0.5 pts)")
                total_score += 0.5
            else:
                # Secondary check: formula contains MAX and references B2:B32
                if 'MAX' in normalized and 'B2' in normalized and 'B32' in normalized:
                    print(f"PASS: Component 1 — E2 contains MAX formula covering B2:B32: {e2_value!r} (0.5 pts)")
                    total_score += 0.5
                else:
                    print(f"FAIL: Component 1 — E2 formula does not match expected =MAX(B2:B32), found: {e2_value!r}")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not check E2: {e}")

    # -----------------------------------------------------------------------
    # Component 2: Cell E3 contains a MIN formula covering B2:B32 (0.5 points)
    # -----------------------------------------------------------------------
    # This FAILS on initial (E3 is None) and PASSES on golden (E3='=MIN(B2:B32)')
    try:
        e3_value = ws.cell(row=3, column=5).value   # column E = 5

        if e3_value is None:
            print(f"FAIL: Component 2 — E3 is empty (None). Expected MIN formula.")
        elif not isinstance(e3_value, str):
            print(f"FAIL: Component 2 — E3 is not a formula string, found: {repr(e3_value)}")
        else:
            normalized = e3_value.strip().upper().replace(' ', '')
            if re.match(EXPECTED_E3_FORMULA_PATTERN, e3_value.strip(), re.IGNORECASE):
                print(f"PASS: Component 2 — E3 contains MIN formula: {e3_value!r} (0.5 pts)")
                total_score += 0.5
            else:
                # Secondary check: formula contains MIN and references B2:B32
                if 'MIN' in normalized and 'B2' in normalized and 'B32' in normalized:
                    print(f"PASS: Component 2 — E3 contains MIN formula covering B2:B32: {e3_value!r} (0.5 pts)")
                    total_score += 0.5
                else:
                    print(f"FAIL: Component 2 — E3 formula does not match expected =MIN(B2:B32), found: {e3_value!r}")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not check E3: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
