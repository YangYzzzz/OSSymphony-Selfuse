"""
Reward Script: Create pivot table in Sheet2 counting POs by supplier and month
Task ID: osworld_calc_pivot_count_invoice_004
Domain: libreoffice_calc
Scoring:
  Component 1: Sheet2 has correct header row with Supplier Name + month labels (0.3 pts)
  Component 2: Correct supplier names as row labels in Sheet2 (0.2 pts)
  Component 3: All pivot COUNT values match expected counts per supplier per month (0.5 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_pivot_count_invoice_004'

# Expected pivot table structure (derived from source data in PurchaseOrders)
# Header: ['Supplier Name', 'Jan', 'Feb', 'Mar', 'Apr', 'May']
EXPECTED_HEADERS = ['Supplier Name', 'Jan', 'Feb', 'Mar', 'Apr', 'May']

# Expected suppliers (sorted for order-independent comparison)
EXPECTED_SUPPLIERS = sorted([
    'Apex Supplies Co.',
    'BlueLine Materials',
    'Horizon Logistics',
    'Zenith Traders',
])

# Expected pivot counts: {supplier -> {month_col_header -> count}}
# Derived from PurchaseOrders source data
EXPECTED_COUNTS = {
    'Apex Supplies Co.':  {'Jan': 2, 'Feb': 2, 'Mar': 1, 'Apr': 2, 'May': 1},
    'BlueLine Materials': {'Jan': 1, 'Feb': 1, 'Mar': 2, 'Apr': 2, 'May': 0},
    'Horizon Logistics':  {'Jan': 0, 'Feb': 1, 'Mar': 2, 'Apr': 1, 'May': 1},
    'Zenith Traders':     {'Jan': 1, 'Feb': 1, 'Mar': 2, 'Apr': 1, 'May': 1},
}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    The task requires Sheet2 to contain a pivot table where:
    - Rows represent suppliers (Supplier Name)
    - Columns represent months (Jan, Feb, Mar, Apr, May)
    - Cell values are COUNT of PO Numbers per supplier per month
    """
    total_score = 0.0

    # Precondition gate: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: Sheet2 must exist
    if 'Sheet2' not in wb.sheetnames:
        print("FAIL: Sheet2 not found in workbook (pivot table sheet missing)")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws2 = wb['Sheet2']

    # Precondition gate: Sheet2 must not be empty
    if ws2.max_row < 2 or ws2.max_column < 2:
        print(f"FAIL: Sheet2 is empty or too small (max_row={ws2.max_row}, max_col={ws2.max_column})")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # --------------------------------------------------------------------------
    # Component 1: Sheet2 has correct header row — 0.3 points
    # Checks: row 1 contains month column headers (Jan through May) and
    #         the first column header is 'Supplier Name' (or similar label)
    # This FAILS on initial (Sheet2 is empty) and PASSES on golden.
    # --------------------------------------------------------------------------
    try:
        # Read the actual header row (row 1)
        header_row = [ws2.cell(row=1, column=c).value for c in range(1, ws2.max_column + 1)]
        # Strip None values from the end
        while header_row and header_row[-1] is None:
            header_row.pop()

        month_headers_found = []
        month_headers_expected = ['Jan', 'Feb', 'Mar', 'Apr', 'May']

        for h in header_row[1:]:  # skip first column (supplier label)
            if h is not None:
                month_headers_found.append(str(h).strip())

        # Check that all expected month headers are present
        months_ok = all(m in month_headers_found for m in month_headers_expected)
        # Check first column has a supplier label
        first_col_label = str(header_row[0]).strip() if header_row[0] else ''
        supplier_label_ok = len(first_col_label) > 0

        if months_ok and supplier_label_ok:
            print(f"PASS: Component 1 — Header row contains month labels {month_headers_found} and supplier column '{first_col_label}' (0.3 pts)")
            total_score += 0.3
        else:
            if not months_ok:
                missing_months = [m for m in month_headers_expected if m not in month_headers_found]
                print(f"FAIL: Component 1 — Header row missing month labels: {missing_months}. Found: {month_headers_found}")
            if not supplier_label_ok:
                print(f"FAIL: Component 1 — First column header is empty or missing.")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # --------------------------------------------------------------------------
    # Component 2: Correct supplier names as row labels — 0.2 points
    # Checks: rows 2+ in column A contain the 4 expected supplier names.
    # This FAILS on initial (Sheet2 is empty) and PASSES on golden.
    # --------------------------------------------------------------------------
    try:
        supplier_col = []
        for r in range(2, ws2.max_row + 1):
            val = ws2.cell(row=r, column=1).value
            if val is not None:
                supplier_col.append(str(val).strip())

        actual_suppliers_sorted = sorted(supplier_col)

        if actual_suppliers_sorted == EXPECTED_SUPPLIERS:
            print(f"PASS: Component 2 — All 4 supplier names present as row labels: {actual_suppliers_sorted} (0.2 pts)")
            total_score += 0.2
        else:
            missing = [s for s in EXPECTED_SUPPLIERS if s not in supplier_col]
            extra = [s for s in supplier_col if s not in EXPECTED_SUPPLIERS]
            print(f"FAIL: Component 2 — Supplier row labels mismatch. Missing: {missing}, Extra: {extra}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # --------------------------------------------------------------------------
    # Component 3: All pivot COUNT values match expected counts — 0.5 points
    # Checks every cell of the pivot table body against the expected counts
    # computed from the PurchaseOrders source data.
    # This FAILS on initial (Sheet2 is empty) and PASSES on golden.
    # --------------------------------------------------------------------------
    try:
        # Build a map of header -> column index from row 1
        header_to_col = {}
        for c in range(1, ws2.max_column + 1):
            h = ws2.cell(row=1, column=c).value
            if h is not None:
                header_to_col[str(h).strip()] = c

        # Build a map of supplier -> row index from column 1
        supplier_to_row = {}
        for r in range(2, ws2.max_row + 1):
            s = ws2.cell(row=r, column=1).value
            if s is not None:
                supplier_to_row[str(s).strip()] = r

        month_labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May']

        # Check that we can look up all expected headers/suppliers
        if not all(m in header_to_col for m in month_labels):
            print("FAIL: Component 3 — Month columns not found in header row, cannot verify counts.")
        elif not all(s in supplier_to_row for s in EXPECTED_SUPPLIERS):
            missing_sup = [s for s in EXPECTED_SUPPLIERS if s not in supplier_to_row]
            print(f"FAIL: Component 3 — Missing suppliers in pivot: {missing_sup}")
        else:
            errors = []
            total_cells = 0
            correct_cells = 0

            for supplier, expected_months in EXPECTED_COUNTS.items():
                row_idx = supplier_to_row[supplier]
                for month, expected_count in expected_months.items():
                    total_cells += 1
                    col_idx = header_to_col[month]
                    actual = ws2.cell(row=row_idx, column=col_idx).value

                    # Count must be numeric and match expected (allow int or float representation)
                    try:
                        actual_int = int(actual) if actual is not None else None
                    except (ValueError, TypeError):
                        actual_int = None

                    if actual_int == expected_count:
                        correct_cells += 1
                    else:
                        errors.append(
                            f"  [{supplier}][{month}]: expected={expected_count}, got={actual!r}"
                        )

            if correct_cells == total_cells:
                print(f"PASS: Component 3 — All {total_cells} pivot count cells match expected values (0.5 pts)")
                total_score += 0.5
            else:
                # Partial credit not awarded for this component (binary: all cells must match)
                print(f"FAIL: Component 3 — {correct_cells}/{total_cells} pivot count cells correct. Mismatches:")
                for err in errors[:10]:  # cap output
                    print(err)
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
