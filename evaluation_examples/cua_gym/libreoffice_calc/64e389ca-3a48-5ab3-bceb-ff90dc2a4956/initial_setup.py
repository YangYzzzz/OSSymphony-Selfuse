"""
Initial Setup: Purchase Order data for pivot table task
Task ID: osworld_calc_pivot_count_invoice_004
Domain: libreoffice_calc

Creates Sheet1 with purchase order data (PO Number, Supplier Name, Order Date,
Item Category, Order Value) and an empty Sheet2 where the agent will create
the pivot table.
"""

import os
import shlex
import subprocess
import time
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_pivot_count_invoice_004'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ---- Sheet 1: PurchaseOrders ----
    ws1 = wb.active
    ws1.title = 'PurchaseOrders'

    # Headers
    headers = ['PO Number', 'Supplier Name', 'Order Date', 'Item Category', 'Order Value']
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font = Font(name='Calibri', bold=True, color='FFFFFFFF', size=11)
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Purchase order data (realistic, 25 rows spread over Jan-May 2025)
    orders = [
        ('PO-2025-001', 'Apex Supplies Co.',    date(2025, 1, 10),  'Electronics',  4200.00),
        ('PO-2025-002', 'BlueLine Materials',   date(2025, 1, 15),  'Office',        890.50),
        ('PO-2025-003', 'Apex Supplies Co.',    date(2025, 1, 22),  'Machinery',    7850.00),
        ('PO-2025-004', 'Zenith Traders',       date(2025, 1, 28),  'Electronics',  3100.00),
        ('PO-2025-005', 'BlueLine Materials',   date(2025, 2,  5),  'Office',       1450.00),
        ('PO-2025-006', 'Apex Supplies Co.',    date(2025, 2, 11),  'Electronics',  5600.00),
        ('PO-2025-007', 'Horizon Logistics',    date(2025, 2, 14),  'Logistics',    2300.00),
        ('PO-2025-008', 'Zenith Traders',       date(2025, 2, 19),  'Electronics',  4750.00),
        ('PO-2025-009', 'Apex Supplies Co.',    date(2025, 2, 25),  'Machinery',    9200.00),
        ('PO-2025-010', 'BlueLine Materials',   date(2025, 3,  3),  'Office',        670.00),
        ('PO-2025-011', 'Horizon Logistics',    date(2025, 3,  8),  'Logistics',    3400.00),
        ('PO-2025-012', 'Zenith Traders',       date(2025, 3, 12),  'Electronics',  6100.00),
        ('PO-2025-013', 'Apex Supplies Co.',    date(2025, 3, 18),  'Electronics',  3800.00),
        ('PO-2025-014', 'BlueLine Materials',   date(2025, 3, 24),  'Machinery',    5200.00),
        ('PO-2025-015', 'Horizon Logistics',    date(2025, 3, 27),  'Logistics',    1950.00),
        ('PO-2025-016', 'Zenith Traders',       date(2025, 3, 31),  'Office',        820.00),
        ('PO-2025-017', 'Apex Supplies Co.',    date(2025, 4,  7),  'Electronics',  7100.00),
        ('PO-2025-018', 'BlueLine Materials',   date(2025, 4, 11),  'Office',       1100.00),
        ('PO-2025-019', 'Horizon Logistics',    date(2025, 4, 16),  'Logistics',    4600.00),
        ('PO-2025-020', 'Zenith Traders',       date(2025, 4, 22),  'Electronics',  8300.00),
        ('PO-2025-021', 'Apex Supplies Co.',    date(2025, 4, 28),  'Machinery',    2700.00),
        ('PO-2025-022', 'BlueLine Materials',   date(2025, 4, 30),  'Electronics',  3500.00),
        ('PO-2025-023', 'Horizon Logistics',    date(2025, 5,  5),  'Logistics',    6200.00),
        ('PO-2025-024', 'Apex Supplies Co.',    date(2025, 5,  9),  'Electronics',  4900.00),
        ('PO-2025-025', 'Zenith Traders',       date(2025, 5, 14),  'Electronics',  5500.00),
    ]

    row_fill_light = PatternFill(start_color='FFDCE6F1', end_color='FFDCE6F1', fill_type='solid')
    data_font = Font(name='Calibri', size=11)

    for r, (po, supplier, od, cat, val) in enumerate(orders, 2):
        fill = row_fill_light if r % 2 == 0 else PatternFill(fill_type=None)
        row_data = [po, supplier, od, cat, val]
        for c, cell_val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=cell_val)
            cell.font = data_font
            cell.border = border
            if r % 2 == 0:
                cell.fill = row_fill_light
            if c == 3:  # Order Date column
                cell.number_format = 'yyyy-mm-dd'
                cell.alignment = Alignment(horizontal='center')
            elif c == 5:  # Order Value column
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')

    # Set column widths for readability
    ws1.column_dimensions['A'].width = 14  # PO Number
    ws1.column_dimensions['B'].width = 22  # Supplier Name
    ws1.column_dimensions['C'].width = 14  # Order Date
    ws1.column_dimensions['D'].width = 18  # Item Category
    ws1.column_dimensions['E'].width = 14  # Order Value
    ws1.freeze_panes = 'A2'

    # ---- Sheet 2: Sheet2 (empty — agent creates pivot here) ----
    ws2 = wb.create_sheet('Sheet2')
    ws2['A1'].value = None  # intentionally empty

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
