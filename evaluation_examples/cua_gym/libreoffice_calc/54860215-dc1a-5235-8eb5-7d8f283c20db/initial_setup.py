"""
Initial Setup: International Orders spreadsheet with country list
Task ID: calc_gcv_069
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_069'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: International Orders ---
    ws1 = wb.active
    ws1.title = 'Sheet1'

    headers = ['Order ID', 'Customer', 'Country']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # 24 rows of realistic order data; Column C (Country) left empty
    orders = [
        ['ORD-2025-001', 'Elena Vasquez'],
        ['ORD-2025-002', 'Rajesh Patel'],
        ['ORD-2025-003', 'Sophie Laurent'],
        ['ORD-2025-004', 'Takeshi Yamamoto'],
        ['ORD-2025-005', 'Maria Gonzalez'],
        ['ORD-2025-006', 'Liam O\'Brien'],
        ['ORD-2025-007', 'Chen Wei'],
        ['ORD-2025-008', 'Anna Kowalski'],
        ['ORD-2025-009', 'James Mitchell'],
        ['ORD-2025-010', 'Fatima Al-Hassan'],
        ['ORD-2025-011', 'Diego Rivera'],
        ['ORD-2025-012', 'Yuki Tanaka'],
        ['ORD-2025-013', 'Priya Sharma'],
        ['ORD-2025-014', 'Lucas Dubois'],
        ['ORD-2025-015', 'Olivia Bennett'],
        ['ORD-2025-016', 'Hans Mueller'],
        ['ORD-2025-017', 'Isabella Rossi'],
        ['ORD-2025-018', 'Arjun Mehta'],
        ['ORD-2025-019', 'Sarah Kim'],
        ['ORD-2025-020', 'Carlos Mendes'],
        ['ORD-2025-021', 'Emma Johansson'],
        ['ORD-2025-022', 'Ahmed Nasser'],
        ['ORD-2025-023', 'Natalie Park'],
        ['ORD-2025-024', 'Marco Bianchi'],
    ]

    for r, row_data in enumerate(orders, 2):
        ws1.cell(row=r, column=1, value=row_data[0])
        ws1.cell(row=r, column=2, value=row_data[1])
        # Column C intentionally left empty - no Country values

    # Set reasonable column widths
    ws1.column_dimensions['A'].width = 16
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 16

    # --- Sheet2: Country List ---
    ws2 = wb.create_sheet('Sheet2')
    countries = [
        'USA', 'Canada', 'UK', 'Germany', 'France',
        'Japan', 'Australia', 'Brazil', 'India', 'Mexico',
        'Italy', 'Spain', 'China', 'South Korea', 'Netherlands'
    ]
    for r, country in enumerate(countries, 1):
        ws2.cell(row=r, column=1, value=country)

    ws2.column_dimensions['A'].width = 16

    # NO named ranges defined
    # NO data validation defined

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
