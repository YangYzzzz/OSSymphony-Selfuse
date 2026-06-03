"""
Initial Setup: Monthly Household Budget Tracker
Task ID: calc_grs_004
Domain: libreoffice_calc

Creates a spreadsheet with budget data (Income & Expense categories) across two sheets.
Initial state has raw data only - NO grouping, NO conditional formatting on variance,
NO subtotal/variance formulas, NO charts.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_004'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Budget Tracker ----
    ws = wb.active
    ws.title = "Budget Tracker"

    # Styling constants
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    title_font = Font(name="Calibri", size=14, bold=True, color="2F5496")
    section_font = Font(name="Calibri", size=11, bold=True, color="2F5496")
    section_fill = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")
    currency_fmt = '$#,##0.00'
    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    # Row 1: Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "Monthly Household Budget Tracker - April 2025"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Row 2: blank separator
    # Row 3: Headers
    headers = ["Category", "Item", "Budgeted", "Actual", "Variance"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # --- Data layout ---
    # We'll build rows with: (category_group, item_name, budgeted, actual)
    # Variance column left EMPTY (task asks agent to create it)

    row = 4

    # ---- INCOME SECTION ----
    ws.cell(row=row, column=1, value="INCOME").font = section_font
    ws.cell(row=row, column=1).fill = section_fill
    for c in range(1, 6):
        ws.cell(row=row, column=c).fill = section_fill
        ws.cell(row=row, column=c).border = thin_border
    income_header_row = row
    row += 1

    income_items = [
        ("Income", "Salary", 6500.00, 6500.00),
        ("Income", "Freelance Work", 1200.00, 1450.00),
        ("Income", "Other Income", 300.00, 175.00),
    ]
    income_start = row
    for cat, item, budgeted, actual in income_items:
        ws.cell(row=row, column=1, value=cat).border = thin_border
        ws.cell(row=row, column=2, value=item).border = thin_border
        c_bud = ws.cell(row=row, column=3, value=budgeted)
        c_bud.number_format = currency_fmt
        c_bud.border = thin_border
        c_act = ws.cell(row=row, column=4, value=actual)
        c_act.number_format = currency_fmt
        c_act.border = thin_border
        # Variance column intentionally empty
        ws.cell(row=row, column=5).border = thin_border
        ws.cell(row=row, column=5).number_format = currency_fmt
        row += 1
    income_end = row - 1

    # Income Subtotal row (values only, no formulas - task asks agent to add formulas)
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2, value="Income Subtotal").font = Font(bold=True)
    ws.cell(row=row, column=2).border = thin_border
    for c in range(3, 6):
        ws.cell(row=row, column=c).border = thin_border
        ws.cell(row=row, column=c).number_format = currency_fmt
    # Leave subtotal cells empty - agent should add SUM formulas
    income_subtotal_row = row
    row += 1

    # Blank separator
    row += 1

    # ---- EXPENSE SECTION ----
    ws.cell(row=row, column=1, value="EXPENSES").font = section_font
    ws.cell(row=row, column=1).fill = section_fill
    for c in range(1, 6):
        ws.cell(row=row, column=c).fill = section_fill
        ws.cell(row=row, column=c).border = thin_border
    expense_header_row = row
    row += 1

    expense_categories = {
        "Housing": [
            ("Mortgage/Rent", 1800.00, 1800.00),
            ("Home Insurance", 150.00, 150.00),
            ("Property Tax", 250.00, 250.00),
            ("Maintenance", 100.00, 235.00),
        ],
        "Transportation": [
            ("Car Payment", 450.00, 450.00),
            ("Gas", 200.00, 267.50),
            ("Car Insurance", 120.00, 120.00),
            ("Parking/Tolls", 50.00, 62.00),
        ],
        "Food": [
            ("Groceries", 600.00, 742.30),
            ("Dining Out", 200.00, 318.75),
            ("Coffee Shops", 60.00, 87.50),
        ],
        "Utilities": [
            ("Electricity", 130.00, 148.20),
            ("Water/Sewer", 65.00, 63.40),
            ("Internet", 80.00, 80.00),
            ("Phone", 120.00, 120.00),
        ],
        "Healthcare": [
            ("Health Insurance", 400.00, 400.00),
            ("Prescriptions", 50.00, 72.50),
            ("Doctor Visits", 75.00, 0.00),
        ],
        "Entertainment": [
            ("Streaming Services", 45.00, 45.00),
            ("Hobbies", 100.00, 156.80),
            ("Events/Outings", 80.00, 125.00),
        ],
        "Savings": [
            ("Emergency Fund", 500.00, 500.00),
            ("Retirement (401k)", 650.00, 650.00),
            ("Vacation Fund", 200.00, 100.00),
        ],
    }

    for cat_name, items in expense_categories.items():
        # Category sub-header
        ws.cell(row=row, column=1, value=cat_name).font = Font(bold=True, italic=True)
        ws.cell(row=row, column=1).fill = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = thin_border
            if c > 1:
                ws.cell(row=row, column=c).fill = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
        cat_header_row = row
        row += 1

        item_start = row
        for item_name, budgeted, actual in items:
            ws.cell(row=row, column=1, value=cat_name).border = thin_border
            ws.cell(row=row, column=2, value=item_name).border = thin_border
            c_bud = ws.cell(row=row, column=3, value=budgeted)
            c_bud.number_format = currency_fmt
            c_bud.border = thin_border
            c_act = ws.cell(row=row, column=4, value=actual)
            c_act.number_format = currency_fmt
            c_act.border = thin_border
            ws.cell(row=row, column=5).border = thin_border
            ws.cell(row=row, column=5).number_format = currency_fmt
            row += 1
        item_end = row - 1

        # Category subtotal row (empty - agent adds formulas)
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2, value=f"{cat_name} Subtotal").font = Font(bold=True)
        ws.cell(row=row, column=2).border = thin_border
        for c in range(3, 6):
            ws.cell(row=row, column=c).border = thin_border
            ws.cell(row=row, column=c).number_format = currency_fmt
        row += 1

    # Blank separator
    row += 1

    # Net Income row (empty - agent adds formula)
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2, value="NET INCOME").font = Font(name="Calibri", size=12, bold=True, color="2F5496")
    ws.cell(row=row, column=2).border = thin_border
    for c in range(3, 6):
        ws.cell(row=row, column=c).border = thin_border
        ws.cell(row=row, column=c).number_format = currency_fmt
        ws.cell(row=row, column=c).font = Font(bold=True)

    # Freeze header row
    ws.freeze_panes = "A4"

    # ---- Sheet 2: Charts ----
    ws2 = wb.create_sheet("Charts")
    ws2["A1"] = "Expense Distribution Data"
    ws2["A1"].font = Font(size=12, bold=True)

    # Summary data for charts (agent will use this to create charts)
    chart_headers = ["Category", "Budgeted", "Actual"]
    for c, h in enumerate(chart_headers, 1):
        cell = ws2.cell(row=3, column=c, value=h)
        cell.font = Font(bold=True)
        cell.border = thin_border

    chart_data = [
        ("Housing", 2300.00, 2435.00),
        ("Transportation", 820.00, 899.50),
        ("Food", 860.00, 1148.55),
        ("Utilities", 395.00, 411.60),
        ("Healthcare", 525.00, 472.50),
        ("Entertainment", 225.00, 326.80),
        ("Savings", 1350.00, 1250.00),
    ]
    for r, (cat, bud, act) in enumerate(chart_data, 4):
        ws2.cell(row=r, column=1, value=cat).border = thin_border
        c_bud = ws2.cell(row=r, column=2, value=bud)
        c_bud.number_format = currency_fmt
        c_bud.border = thin_border
        c_act = ws2.cell(row=r, column=3, value=act)
        c_act.number_format = currency_fmt
        c_act.border = thin_border

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14

    # NO charts added - task asks agent to create them

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
