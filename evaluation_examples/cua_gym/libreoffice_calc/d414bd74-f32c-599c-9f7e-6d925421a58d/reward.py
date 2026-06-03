"""
Reward Script: Monthly Household Budget Tracker
Task ID: calc_grs_004
Domain: libreoffice_calc
Scoring:
  Component 1: Variance formulas in column E (0.25)
  Component 2: SUM subtotal formulas in C and D columns (0.20)
  Component 3: NET INCOME formulas in row 50 (0.10)
  Component 4: Conditional formatting on variance column (0.15)
  Component 5: Row grouping/outline for detail rows (0.15)
  Component 6: Charts on Charts sheet (pie + bar) (0.15)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_004'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Budget Tracker' sheet must exist
    if 'Budget Tracker' not in wb.sheetnames:
        print("FAIL: 'Budget Tracker' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Budget Tracker']

    # ---------------------------------------------------------------
    # Component 1: Variance formulas in column E (0.25 points)
    # The golden file has =Dn-Cn formulas in E for each data row and subtotal row.
    # Initial file has NO values in column E at all.
    # ---------------------------------------------------------------
    try:
        # Check variance formulas in data rows
        # Data rows: 5-7 (income), 12-15 (housing), 18-21 (transport),
        # 24-26 (food), 29-32 (utilities), 35-37 (healthcare),
        # 40-42 (entertainment), 45-47 (savings)
        data_rows = [5, 6, 7, 12, 13, 14, 15, 18, 19, 20, 21,
                     24, 25, 26, 29, 30, 31, 32, 35, 36, 37,
                     40, 41, 42, 45, 46, 47]
        subtotal_rows = [8, 16, 22, 27, 33, 38, 43, 48]
        all_variance_rows = data_rows + subtotal_rows

        variance_count = 0
        for r in all_variance_rows:
            val = ws.cell(row=r, column=5).value
            if val is not None and isinstance(val, str) and '=' in val:
                variance_count += 1

        variance_ratio = variance_count / len(all_variance_rows)
        if variance_ratio >= 0.8:
            print(f"PASS: Component 1 — Variance formulas found in {variance_count}/{len(all_variance_rows)} rows (0.25 pts)")
            total_score += 0.25
        elif variance_ratio >= 0.5:
            partial = 0.25 * (variance_ratio / 0.8)
            print(f"PARTIAL: Component 1 — Variance formulas in {variance_count}/{len(all_variance_rows)} rows ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {variance_count}/{len(all_variance_rows)} variance formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ---------------------------------------------------------------
    # Component 2: SUM subtotal formulas in C and D columns (0.20 points)
    # Golden has =SUM(...) in C8,D8, C16,D16, C22,D22, C27,D27,
    # C33,D33, C38,D38, C43,D43, C48,D48
    # Initial has None in all these cells.
    # ---------------------------------------------------------------
    try:
        subtotal_cells = [
            ('C8', 'D8'), ('C16', 'D16'), ('C22', 'D22'), ('C27', 'D27'),
            ('C33', 'D33'), ('C38', 'D38'), ('C43', 'D43'), ('C48', 'D48')
        ]
        sum_count = 0
        total_checks = 0
        for c_cell, d_cell in subtotal_cells:
            for cell_ref in [c_cell, d_cell]:
                total_checks += 1
                val = ws[cell_ref].value
                if val is not None and isinstance(val, str) and 'SUM' in val.upper():
                    sum_count += 1

        sum_ratio = sum_count / total_checks
        if sum_ratio >= 0.8:
            print(f"PASS: Component 2 — SUM subtotal formulas found in {sum_count}/{total_checks} cells (0.20 pts)")
            total_score += 0.20
        elif sum_ratio >= 0.4:
            partial = 0.20 * (sum_ratio / 0.8)
            print(f"PARTIAL: Component 2 — SUM formulas in {sum_count}/{total_checks} cells ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {sum_count}/{total_checks} SUM subtotal formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---------------------------------------------------------------
    # Component 3: NET INCOME formulas in row 50 (0.10 points)
    # Golden has formulas in C50, D50, E50 that compute net income.
    # Initial has None in these cells.
    # ---------------------------------------------------------------
    try:
        net_income_count = 0
        for col in [3, 4, 5]:  # C, D, E
            val = ws.cell(row=50, column=col).value
            if val is not None and isinstance(val, str) and '=' in val:
                net_income_count += 1

        if net_income_count >= 2:
            print(f"PASS: Component 3 — NET INCOME formulas found ({net_income_count}/3 cells have formulas) (0.10 pts)")
            total_score += 0.10
        elif net_income_count >= 1:
            print(f"PARTIAL: Component 3 — Only {net_income_count}/3 NET INCOME formulas (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 3 — No NET INCOME formulas in row 50")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ---------------------------------------------------------------
    # Component 4: Conditional formatting on variance column (0.15 points)
    # Golden has 2 rules on E5:E50: greaterThan 0 (green) and lessThan 0 (red).
    # Initial has NO conditional formatting.
    # ---------------------------------------------------------------
    try:
        cf_rules_on_e = []
        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            # Check if any conditional formatting range includes column E
            if 'E' in cf_range:
                for rule in cf.rules:
                    cf_rules_on_e.append(rule)

        if len(cf_rules_on_e) >= 2:
            # Check that we have both positive (green) and negative (red) rules
            has_positive_rule = False
            has_negative_rule = False
            for rule in cf_rules_on_e:
                op = getattr(rule, 'operator', None)
                if op in ('greaterThan', 'greaterThanOrEqual'):
                    has_positive_rule = True
                elif op in ('lessThan', 'lessThanOrEqual'):
                    has_negative_rule = True

            if has_positive_rule and has_negative_rule:
                print(f"PASS: Component 4 — Conditional formatting with positive/negative variance rules (0.15 pts)")
                total_score += 0.15
            else:
                print(f"PARTIAL: Component 4 — CF rules found but missing positive/negative distinction (0.08 pts)")
                total_score += 0.08
        elif len(cf_rules_on_e) >= 1:
            print(f"PARTIAL: Component 4 — Only {len(cf_rules_on_e)} CF rule(s) found on E column (0.07 pts)")
            total_score += 0.07
        else:
            print(f"FAIL: Component 4 — No conditional formatting on variance column E")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ---------------------------------------------------------------
    # Component 5: Row grouping/outline for detail rows (0.15 points)
    # Golden has outline_level=1 on detail rows (items within each category).
    # Initial has NO row grouping at all.
    # ---------------------------------------------------------------
    try:
        # Expected grouped rows: detail rows within each category
        expected_grouped = [
            5, 6, 7,           # Income items
            12, 13, 14, 15,    # Housing items
            18, 19, 20, 21,    # Transportation items
            24, 25, 26,        # Food items
            29, 30, 31, 32,    # Utilities items
            35, 36, 37,        # Healthcare items
            40, 41, 42,        # Entertainment items
            45, 46, 47,        # Savings items
        ]

        grouped_count = 0
        for r in expected_grouped:
            rd = ws.row_dimensions.get(r)
            if rd and rd.outline_level and rd.outline_level >= 1:
                grouped_count += 1

        group_ratio = grouped_count / len(expected_grouped)
        if group_ratio >= 0.7:
            print(f"PASS: Component 5 — Row grouping found on {grouped_count}/{len(expected_grouped)} detail rows (0.15 pts)")
            total_score += 0.15
        elif group_ratio >= 0.3:
            partial = 0.15 * (group_ratio / 0.7)
            print(f"PARTIAL: Component 5 — Row grouping on {grouped_count}/{len(expected_grouped)} rows ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — Only {grouped_count}/{len(expected_grouped)} rows have grouping")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # ---------------------------------------------------------------
    # Component 6: Charts on Charts sheet (0.15 points)
    # Golden has 2 charts: PieChart (expense distribution) and BarChart (budgeted vs actual).
    # Initial has 0 charts.
    # ---------------------------------------------------------------
    try:
        if 'Charts' not in wb.sheetnames:
            print("FAIL: Component 6 — 'Charts' sheet not found")
        else:
            ws_charts = wb['Charts']
            charts = ws_charts._charts
            num_charts = len(charts)

            if num_charts >= 2:
                # Check chart types
                chart_types = [type(c).__name__ for c in charts]
                has_pie = any('Pie' in ct for ct in chart_types)
                has_bar = any('Bar' in ct for ct in chart_types)

                if has_pie and has_bar:
                    print(f"PASS: Component 6 — Found PieChart and BarChart on Charts sheet (0.15 pts)")
                    total_score += 0.15
                else:
                    print(f"PARTIAL: Component 6 — Found {num_charts} chart(s) but types={chart_types}, expected Pie+Bar (0.10 pts)")
                    total_score += 0.10
            elif num_charts == 1:
                print(f"PARTIAL: Component 6 — Only 1 chart found, expected 2 (pie + bar) (0.07 pts)")
                total_score += 0.07
            else:
                print(f"FAIL: Component 6 — No charts found on Charts sheet")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
