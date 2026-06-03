"""
Reward Script: Q4 Performance Review — IFS Rating Formula + Conditional Formatting
Task ID: calc_hr_performance_rating_004
Domain: libreoffice_calc

Scoring Rubric:
  Component 1a: E2 contains correct IFS formula (0.25 pts)
  Component 1b: All 53 cells E2:E54 contain IFS formulas (0.25 pts)
  Component 2a: Green CF rule ($D>=3.5, background #70AD47) on A2:F54 (0.20 pts)
  Component 2b: Yellow CF rule (AND $D>=2.5 $D<3.5, background #FFFF00) on A2:F54 (0.15 pts)
  Component 2c: Red CF rule ($D<2.5, background #FF0000) on A2:F54 (0.15 pts)
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_performance_rating_004'
SHEET_NAME = 'Q4 Reviews'

# Expected IFS formula pattern: row reference varies per row (D2, D3, ...)
IFS_PATTERN = re.compile(
    r'=IFS\(\s*D\d+\s*>=\s*4\.5\s*,\s*"Exceptional"\s*,\s*D\d+\s*>=\s*3\.5\s*,\s*"Exceeds"\s*,'
    r'\s*D\d+\s*>=\s*2\.5\s*,\s*"Meets"\s*,\s*D\d+\s*>=\s*1\.5\s*,\s*"Below"\s*,\s*TRUE\s*,\s*"Unsatisfactory"\s*\)',
    re.IGNORECASE
)


def find_cf_rule_with_color(rules, color_substring):
    """Find the first CF rule whose dxf fill color contains the given substring."""
    for rule in rules:
        try:
            if rule.dxf and rule.dxf.fill and rule.dxf.fill.fgColor:
                fg = rule.dxf.fill.fgColor.rgb
                if fg and color_substring.upper() in fg.upper():
                    return rule
        except Exception:
            pass
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition gate: load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: sheet must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # -----------------------------------------------------------------------
    # Component 1a: E2 contains a correct IFS formula (0.25 pts)
    # FAILS on initial (E2 is None) -> PASSES on golden (IFS formula present)
    # -----------------------------------------------------------------------
    try:
        e2_val = ws.cell(row=2, column=5).value
        e2_has_ifs = (
            e2_val is not None
            and isinstance(e2_val, str)
            and IFS_PATTERN.match(e2_val.strip()) is not None
        )
        if e2_has_ifs:
            print(f"PASS: Component 1a — E2 contains correct IFS formula ({e2_val[:70]}...) (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1a — E2 expected IFS formula, found: {repr(e2_val)}")
    except Exception as e:
        print(f"ERROR: Component 1a — {e}")

    # -----------------------------------------------------------------------
    # Component 1b: All 53 cells E2:E54 contain IFS formulas (0.25 pts)
    # FAILS on initial (all None) -> PASSES on golden (all 53 cells populated)
    # -----------------------------------------------------------------------
    try:
        formula_count = sum(
            1 for row in range(2, 55)
            if (v := ws.cell(row=row, column=5).value) is not None
            and isinstance(v, str)
            and IFS_PATTERN.match(v.strip()) is not None
        )
        if formula_count == 53:
            print(f"PASS: Component 1b — All 53 cells E2:E54 contain correct IFS formulas (0.25 pts)")
            total_score += 0.25
        else:
            missing = [
                f"E{row}"
                for row in range(2, 55)
                if not (
                    (v := ws.cell(row=row, column=5).value) is not None
                    and isinstance(v, str)
                    and IFS_PATTERN.match(v.strip()) is not None
                )
            ]
            print(f"FAIL: Component 1b — Expected 53 IFS formulas, found {formula_count}. "
                  f"Missing/wrong: {missing[:5]}")
    except Exception as e:
        print(f"ERROR: Component 1b — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Conditional formatting on A2:F54 (3 rules: green/yellow/red)
    # FAILS on initial (no CF at all) -> PASSES on golden (3 rules present)
    # -----------------------------------------------------------------------
    try:
        cf_list = list(ws.conditional_formatting)
        target_cf = None
        for cf in cf_list:
            if 'A2' in str(cf) and 'F54' in str(cf):
                target_cf = cf
                break

        if target_cf is None:
            print(f"FAIL: Component 2 — No conditional formatting found on A2:F54 "
                  f"(total CF ranges: {len(cf_list)})")
        else:
            rules = target_cf.rules
            print(f"INFO: Found CF on A2:F54 with {len(rules)} rules")

            # Component 2a: Green rule ($D>=3.5, background #70AD47) (0.20 pts)
            try:
                green_rule = find_cf_rule_with_color(rules, '70AD47')
                formula_str = ' '.join(green_rule.formula) if green_rule and green_rule.formula else ''
                has_gte_3_5 = re.search(r'\$?D\d*\s*>=\s*3\.5', formula_str, re.IGNORECASE) is not None
                if green_rule and has_gte_3_5:
                    fg = green_rule.dxf.fill.fgColor.rgb
                    print(f"PASS: Component 2a — Green rule ($D>=3.5, #{fg}) found: formula={formula_str} (0.20 pts)")
                    total_score += 0.20
                elif green_rule:
                    print(f"FAIL: Component 2a — Green color found but formula wrong: {formula_str}")
                else:
                    print(f"FAIL: Component 2a — Green rule ($D>=3.5, #70AD47) not found")
            except Exception as e:
                print(f"ERROR: Component 2a — {e}")

            # Component 2b: Yellow rule (AND $D>=2.5 $D<3.5, background #FFFF00) (0.15 pts)
            try:
                yellow_rule = find_cf_rule_with_color(rules, 'FFFF00')
                formula_str = ' '.join(yellow_rule.formula) if yellow_rule and yellow_rule.formula else ''
                has_and_condition = (
                    re.search(r'AND', formula_str, re.IGNORECASE) is not None
                    and re.search(r'>=\s*2\.5', formula_str) is not None
                    and re.search(r'<\s*3\.5', formula_str) is not None
                )
                if yellow_rule and has_and_condition:
                    fg = yellow_rule.dxf.fill.fgColor.rgb
                    print(f"PASS: Component 2b — Yellow rule (AND $D>=2.5 $D<3.5, #{fg}) found: formula={formula_str} (0.15 pts)")
                    total_score += 0.15
                elif yellow_rule:
                    print(f"FAIL: Component 2b — Yellow color found but formula wrong: {formula_str}")
                else:
                    print(f"FAIL: Component 2b — Yellow rule (AND $D>=2.5 $D<3.5, #FFFF00) not found")
            except Exception as e:
                print(f"ERROR: Component 2b — {e}")

            # Component 2c: Red rule ($D<2.5, background #FF0000) (0.15 pts)
            try:
                red_rule = find_cf_rule_with_color(rules, 'FF0000')
                formula_str = ' '.join(red_rule.formula) if red_rule and red_rule.formula else ''
                has_lt_2_5 = re.search(r'\$?D\d*\s*<\s*2\.5', formula_str, re.IGNORECASE) is not None
                if red_rule and has_lt_2_5:
                    fg = red_rule.dxf.fill.fgColor.rgb
                    print(f"PASS: Component 2c — Red rule ($D<2.5, #{fg}) found: formula={formula_str} (0.15 pts)")
                    total_score += 0.15
                elif red_rule:
                    print(f"FAIL: Component 2c — Red color found but formula wrong: {formula_str}")
                else:
                    print(f"FAIL: Component 2c — Red rule ($D<2.5, #FF0000) not found")
            except Exception as e:
                print(f"ERROR: Component 2c — {e}")

    except Exception as e:
        print(f"ERROR: Component 2 (CF check) — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
