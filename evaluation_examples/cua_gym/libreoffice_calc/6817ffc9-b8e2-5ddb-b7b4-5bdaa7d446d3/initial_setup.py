"""
Initial Setup: Q4 Performance Review Sheet
Task ID: calc_hr_performance_rating_004
Domain: libreoffice_calc
Creates Q4 Reviews sheet with employee data, empty Rating Label column, no conditional formatting.
"""

import os
import openpyxl
from datetime import date, timedelta
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_performance_rating_004'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Q4 Reviews'

    # Headers
    ws['A1'] = 'Emp ID'
    ws['B1'] = 'Name'
    ws['C1'] = 'Manager'
    ws['D1'] = 'Score'
    ws['E1'] = 'Rating Label'
    ws['F1'] = 'Review Date'

    # Realistic employee data (53 rows: rows 2-54)
    employees = [
        ('E001', 'Sarah Chen',        'Michael Torres',   4.8, '2024-10-05'),
        ('E002', 'Marcus Johnson',     'Linda Park',       3.2, '2024-10-07'),
        ('E003', 'Priya Patel',        'Michael Torres',   4.5, '2024-10-08'),
        ('E004', 'James O\'Brien',     'Linda Park',       2.1, '2024-10-10'),
        ('E005', 'Natalie Rivera',     'David Kim',        3.7, '2024-10-11'),
        ('E006', 'Kevin Okafor',       'David Kim',        1.8, '2024-10-14'),
        ('E007', 'Amelia Foster',      'Rachel Nguyen',    4.2, '2024-10-15'),
        ('E008', 'Daniel Alvarez',     'Rachel Nguyen',    3.5, '2024-10-16'),
        ('E009', 'Sophia Williams',    'Michael Torres',   2.6, '2024-10-17'),
        ('E010', 'Liam Nakamura',      'Linda Park',       4.9, '2024-10-18'),
        ('E011', 'Isabella Scott',     'David Kim',        1.3, '2024-10-21'),
        ('E012', 'Ethan Brooks',       'Rachel Nguyen',    3.8, '2024-10-22'),
        ('E013', 'Olivia Martinez',    'Michael Torres',   4.0, '2024-10-23'),
        ('E014', 'Noah Thompson',      'Linda Park',       2.4, '2024-10-24'),
        ('E015', 'Emma Davis',         'David Kim',        3.1, '2024-10-25'),
        ('E016', 'Aiden Wilson',       'Rachel Nguyen',    4.6, '2024-10-28'),
        ('E017', 'Charlotte Harris',   'Michael Torres',   1.5, '2024-10-29'),
        ('E018', 'Mason Clark',        'Linda Park',       3.3, '2024-10-30'),
        ('E019', 'Ava Lewis',          'David Kim',        4.1, '2024-11-01'),
        ('E020', 'Logan Robinson',     'Rachel Nguyen',    2.8, '2024-11-04'),
        ('E021', 'Mia Walker',         'Michael Torres',   3.9, '2024-11-05'),
        ('E022', 'Lucas Hall',         'Linda Park',       4.7, '2024-11-06'),
        ('E023', 'Harper Allen',       'David Kim',        1.9, '2024-11-07'),
        ('E024', 'Jackson Young',      'Rachel Nguyen',    3.4, '2024-11-08'),
        ('E025', 'Ella Hernandez',     'Michael Torres',   2.3, '2024-11-11'),
        ('E026', 'Sebastian King',     'Linda Park',       4.4, '2024-11-12'),
        ('E027', 'Scarlett Wright',    'David Kim',        3.6, '2024-11-13'),
        ('E028', 'Henry Lopez',        'Rachel Nguyen',    2.0, '2024-11-14'),
        ('E029', 'Lily Hill',          'Michael Torres',   4.3, '2024-11-15'),
        ('E030', 'Jack Scott',         'Linda Park',       3.0, '2024-11-18'),
        ('E031', 'Zoe Green',          'David Kim',        1.6, '2024-11-19'),
        ('E032', 'Owen Adams',         'Rachel Nguyen',    3.7, '2024-11-20'),
        ('E033', 'Chloe Baker',        'Michael Torres',   4.5, '2024-11-21'),
        ('E034', 'Elijah Gonzalez',    'Linda Park',       2.7, '2024-11-22'),
        ('E035', 'Penelope Nelson',    'David Kim',        3.2, '2024-11-25'),
        ('E036', 'Wyatt Carter',       'Rachel Nguyen',    4.8, '2024-11-26'),
        ('E037', 'Layla Mitchell',     'Michael Torres',   1.4, '2024-11-27'),
        ('E038', 'Gabriel Perez',      'Linda Park',       3.5, '2024-12-02'),
        ('E039', 'Riley Roberts',      'David Kim',        2.9, '2024-12-03'),
        ('E040', 'Julian Turner',      'Rachel Nguyen',    4.0, '2024-12-04'),
        ('E041', 'Zoey Phillips',      'Michael Torres',   3.3, '2024-12-05'),
        ('E042', 'Levi Campbell',      'Linda Park',       1.7, '2024-12-06'),
        ('E043', 'Nora Parker',        'David Kim',        4.6, '2024-12-09'),
        ('E044', 'Isaac Evans',        'Rachel Nguyen',    2.5, '2024-12-10'),
        ('E045', 'Aurora Edwards',     'Michael Torres',   3.8, '2024-12-11'),
        ('E046', 'Elias Collins',      'Linda Park',       4.2, '2024-12-12'),
        ('E047', 'Hannah Stewart',     'David Kim',        2.2, '2024-12-13'),
        ('E048', 'Anthony Sanchez',    'Rachel Nguyen',    3.6, '2024-12-16'),
        ('E049', 'Addison Morris',     'Michael Torres',   1.2, '2024-12-17'),
        ('E050', 'Dominic Rogers',     'Linda Park',       4.9, '2024-12-18'),
        ('E051', 'Brooklyn Reed',      'David Kim',        3.1, '2024-12-19'),
        ('E052', 'Jonathan Cook',      'Rachel Nguyen',    2.6, '2024-12-20'),
        ('E053', 'Valentina Morgan',   'Michael Torres',   4.4, '2024-12-23'),
    ]

    for i, (emp_id, name, manager, score, review_date) in enumerate(employees, 2):
        ws.cell(row=i, column=1, value=emp_id)
        ws.cell(row=i, column=2, value=name)
        ws.cell(row=i, column=3, value=manager)
        ws.cell(row=i, column=4, value=score)
        # Column E (Rating Label) intentionally left empty
        ws.cell(row=i, column=6, value=review_date)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Q4 Reviews')
    print(f'  Rows: 2-54 (53 employee records)')
    print(f'  Column E (Rating Label): empty (task-required)')
    print(f'  Conditional formatting: none (task-required)')

create_initial()
