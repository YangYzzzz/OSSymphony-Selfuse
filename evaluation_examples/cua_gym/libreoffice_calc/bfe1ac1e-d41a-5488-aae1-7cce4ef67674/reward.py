"""
Reward Script: Create Sheet2 summary with monthly totals by product category
Task ID: osworld_calc_sheet2_summary_table_008
Domain: libreoffice_calc

Scoring:
  Component 1 (0.35): Summary sheet exists and has correct table structure
                       (13 rows, 11 columns, header row with Month + 5 categories + 5 pct headers)
  Component 2 (0.35): All 12 data rows have SUMIFS formulas in revenue columns (B-F)
                       referencing the Orders sheet
  Component 3 (0.30): All 12 data rows have percentage formulas in columns G-K
                       that compute each category's share of the monthly total
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_sheet2_summary_table_008'

EXPECTED_MONTHS = [
    'January', 'February', 'March', 'April', 'May', 'June',
    'July', 'August', 'September', 'October', 'November', 'December'
]

EXPECTED_CATEGORIES = ['Electronics', 'Clothing', 'Home & Garden', 'Books', 'Sports']

def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: Orders sheet must exist with data
    if 'Orders' not in wb.sheetnames:
        print("CRITICAL: Orders sheet not found — file may be corrupted")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Summary sheet exists and has correct table structure (0.35 points)
    # This FAILS on initial (empty Summary sheet) and PASSES on golden (13 rows, 11 cols)
    try:
        if 'Summary' not in wb.sheetnames:
            print("FAIL: Component 1 — Summary sheet does not exist")
        else:
            ws = wb['Summary']
            max_row = ws.max_row
            max_col = ws.max_column

            # Check dimensions: should have 13 rows (1 header + 12 months) and 11 columns
            if max_row < 13:
                print(f"FAIL: Component 1 — Summary sheet has only {max_row} rows, expected at least 13 (1 header + 12 months)")
            elif max_col < 11:
                print(f"FAIL: Component 1 — Summary sheet has only {max_col} columns, expected at least 11 (Month + 5 categories + 5 pct columns)")
            else:
                # Check header row has Month label and category names
                header_row = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
                header_str = [str(h).strip() if h is not None else '' for h in header_row]

                has_month_header = any('month' in h.lower() for h in header_str if h)
                has_category_headers = sum(
                    1 for cat in EXPECTED_CATEGORIES
                    if any(cat.lower() in h.lower() for h in header_str if h)
                )

                # Check data rows have month names in column A
                month_names_found = []
                for row_idx in range(2, 14):
                    cell_val = ws.cell(row=row_idx, column=1).value
                    if cell_val and str(cell_val).strip() in EXPECTED_MONTHS:
                        month_names_found.append(str(cell_val).strip())

                if not has_month_header:
                    print(f"FAIL: Component 1 — Header row A1 does not contain 'Month' label. Found: {header_row[0]}")
                elif has_category_headers < 5:
                    print(f"FAIL: Component 1 — Header row missing category names. Found {has_category_headers}/5 categories in: {header_str}")
                elif len(month_names_found) < 12:
                    print(f"FAIL: Component 1 — Only {len(month_names_found)}/12 month names found in column A (rows 2-13). Found: {month_names_found}")
                else:
                    print(f"PASS: Component 1 — Summary sheet has correct structure: {max_row} rows x {max_col} cols, "
                          f"header has Month + {has_category_headers} categories, all 12 months in column A (0.35 pts)")
                    total_score += 0.35
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: SUMIFS formulas in revenue columns B-F for all 12 months (0.35 points)
    # This FAILS on initial (empty cells) and PASSES on golden (SUMIFS formulas referencing Orders)
    try:
        if 'Summary' not in wb.sheetnames:
            print("FAIL: Component 2 — Summary sheet does not exist")
        else:
            ws = wb['Summary']
            sumifs_formula_count = 0
            total_revenue_cells = 0
            cells_with_orders_ref = 0

            for row_idx in range(2, 14):  # 12 data rows
                for col_idx in range(2, 7):  # columns B-F (5 categories)
                    total_revenue_cells += 1
                    cell_val = ws.cell(row=row_idx, column=col_idx).value
                    if cell_val is not None and isinstance(cell_val, str):
                        formula_upper = cell_val.upper().replace(' ', '')
                        if 'SUMIFS' in formula_upper:
                            sumifs_formula_count += 1
                            # Check it references Orders sheet
                            if 'ORDERS' in formula_upper or 'ORDERS!' in formula_upper.replace('!', '!'):
                                cells_with_orders_ref += 1

            if sumifs_formula_count == 0:
                print(f"FAIL: Component 2 — No SUMIFS formulas found in revenue columns B-F (rows 2-13). "
                      f"All {total_revenue_cells} cells appear empty or non-formula")
            elif sumifs_formula_count < total_revenue_cells:
                print(f"FAIL: Component 2 — Only {sumifs_formula_count}/{total_revenue_cells} revenue cells have SUMIFS formulas")
            elif cells_with_orders_ref < sumifs_formula_count:
                print(f"FAIL: Component 2 — SUMIFS formulas found but only {cells_with_orders_ref}/{sumifs_formula_count} reference the Orders sheet")
            else:
                print(f"PASS: Component 2 — All {sumifs_formula_count}/{total_revenue_cells} revenue cells have SUMIFS formulas "
                      f"referencing Orders sheet (0.35 pts)")
                total_score += 0.35
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Percentage formulas in columns G-K for all 12 months (0.30 points)
    # Each category's percentage = category_value / sum_of_all_categories_in_that_month
    # This FAILS on initial (empty cells) and PASSES on golden (division formulas dividing by row sum)
    try:
        if 'Summary' not in wb.sheetnames:
            print("FAIL: Component 3 — Summary sheet does not exist")
        else:
            ws = wb['Summary']
            pct_formula_count = 0
            total_pct_cells = 0
            pct_cells_with_division = 0

            for row_idx in range(2, 14):  # 12 data rows
                for col_idx in range(7, 12):  # columns G-K (5 percentage columns)
                    total_pct_cells += 1
                    cell_val = ws.cell(row=row_idx, column=col_idx).value
                    if cell_val is not None and isinstance(cell_val, str) and cell_val.startswith('='):
                        pct_formula_count += 1
                        # Check it performs division (percentage calculation)
                        # Should divide a revenue cell by sum of revenue cells in that row
                        formula_upper = cell_val.upper().replace(' ', '')
                        if '/' in formula_upper:
                            pct_cells_with_division += 1

            if pct_formula_count == 0:
                print(f"FAIL: Component 3 — No percentage formulas found in columns G-K (rows 2-13). "
                      f"All {total_pct_cells} cells appear empty or non-formula")
            elif pct_formula_count < total_pct_cells:
                print(f"FAIL: Component 3 — Only {pct_formula_count}/{total_pct_cells} percentage cells have formulas")
            elif pct_cells_with_division < pct_formula_count:
                print(f"FAIL: Component 3 — Only {pct_cells_with_division}/{pct_formula_count} percentage formulas use division")
            else:
                print(f"PASS: Component 3 — All {pct_formula_count}/{total_pct_cells} percentage cells have division formulas "
                      f"computing category share of monthly total (0.30 pts)")
                total_score += 0.30
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the VM environment
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
