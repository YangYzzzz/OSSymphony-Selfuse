"""
Initial Setup: E-commerce orders spreadsheet for Sheet2 summary table task
Task ID: osworld_calc_sheet2_summary_table_008
Domain: libreoffice_calc

Creates Sheet1 with e-commerce order data (Order Date, Product Category, Units, Revenue)
and an empty Sheet2 (Summary) that the agent must populate with SUMIFS-based summary.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_sheet2_summary_table_008'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Orders ---
    ws1 = wb.active
    ws1.title = 'Orders'

    # Headers
    headers = ['Order Date', 'Product Category', 'Units', 'Revenue']
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Realistic e-commerce order data
    # 5 product categories: Electronics, Clothing, Home & Garden, Books, Sports
    # Data covering Jan 2024 - Dec 2024 (multiple orders per month per category)
    orders_data = [
        # January 2024
        ('2024-01-03', 'Electronics',   12,  8940.00),
        ('2024-01-05', 'Clothing',       8,  1240.00),
        ('2024-01-07', 'Home & Garden',  5,   875.00),
        ('2024-01-09', 'Books',         15,   449.85),
        ('2024-01-11', 'Sports',         6,  1380.00),
        ('2024-01-14', 'Electronics',   10,  7450.00),
        ('2024-01-16', 'Clothing',      14,  2184.00),
        ('2024-01-18', 'Home & Garden',  9,  1530.00),
        ('2024-01-20', 'Books',         22,   659.78),
        ('2024-01-22', 'Sports',        11,  2530.00),
        ('2024-01-25', 'Electronics',    7,  5215.00),
        ('2024-01-28', 'Clothing',       6,   936.00),
        ('2024-01-30', 'Home & Garden',  3,   510.00),
        # February 2024
        ('2024-02-02', 'Electronics',   15, 11175.00),
        ('2024-02-04', 'Clothing',      10,  1560.00),
        ('2024-02-06', 'Home & Garden',  7,  1190.00),
        ('2024-02-08', 'Books',         18,   539.82),
        ('2024-02-10', 'Sports',         9,  2070.00),
        ('2024-02-13', 'Electronics',    8,  5960.00),
        ('2024-02-15', 'Clothing',      12,  1872.00),
        ('2024-02-17', 'Home & Garden',  4,   680.00),
        ('2024-02-19', 'Books',         25,   749.75),
        ('2024-02-22', 'Sports',         6,  1380.00),
        ('2024-02-25', 'Electronics',   20, 14900.00),
        ('2024-02-27', 'Clothing',       5,   780.00),
        # March 2024
        ('2024-03-01', 'Electronics',   11,  8195.00),
        ('2024-03-03', 'Clothing',       9,  1404.00),
        ('2024-03-05', 'Home & Garden', 12,  2040.00),
        ('2024-03-07', 'Books',         30,   899.70),
        ('2024-03-10', 'Sports',        14,  3220.00),
        ('2024-03-12', 'Electronics',   16, 11920.00),
        ('2024-03-15', 'Clothing',      20,  3120.00),
        ('2024-03-18', 'Home & Garden',  8,  1360.00),
        ('2024-03-20', 'Books',         12,   359.88),
        ('2024-03-22', 'Sports',         7,  1610.00),
        ('2024-03-25', 'Electronics',    5,  3725.00),
        ('2024-03-28', 'Clothing',      11,  1716.00),
        # April 2024
        ('2024-04-02', 'Electronics',   13,  9685.00),
        ('2024-04-04', 'Clothing',      15,  2340.00),
        ('2024-04-06', 'Home & Garden',  6,  1020.00),
        ('2024-04-08', 'Books',         20,   599.80),
        ('2024-04-10', 'Sports',        10,  2300.00),
        ('2024-04-13', 'Electronics',    9,  6705.00),
        ('2024-04-15', 'Clothing',       7,  1092.00),
        ('2024-04-18', 'Home & Garden', 11,  1870.00),
        ('2024-04-20', 'Books',         16,   479.84),
        ('2024-04-23', 'Sports',         8,  1840.00),
        ('2024-04-26', 'Electronics',   18, 13410.00),
        ('2024-04-29', 'Clothing',       4,   624.00),
        # May 2024
        ('2024-05-02', 'Electronics',   14, 10430.00),
        ('2024-05-04', 'Clothing',      11,  1716.00),
        ('2024-05-06', 'Home & Garden', 10,  1700.00),
        ('2024-05-08', 'Books',         24,   719.76),
        ('2024-05-10', 'Sports',        12,  2760.00),
        ('2024-05-13', 'Electronics',    6,  4470.00),
        ('2024-05-15', 'Clothing',      18,  2808.00),
        ('2024-05-18', 'Home & Garden',  5,   850.00),
        ('2024-05-20', 'Books',         10,   299.90),
        ('2024-05-23', 'Sports',        15,  3450.00),
        ('2024-05-26', 'Electronics',   22, 16390.00),
        ('2024-05-29', 'Clothing',       9,  1404.00),
        # June 2024
        ('2024-06-03', 'Electronics',   10,  7450.00),
        ('2024-06-05', 'Clothing',      13,  2028.00),
        ('2024-06-07', 'Home & Garden', 15,  2550.00),
        ('2024-06-09', 'Books',         28,   839.72),
        ('2024-06-11', 'Sports',        20,  4600.00),
        ('2024-06-14', 'Electronics',   17, 12665.00),
        ('2024-06-16', 'Clothing',       6,   936.00),
        ('2024-06-18', 'Home & Garden',  9,  1530.00),
        ('2024-06-20', 'Books',         14,   419.86),
        ('2024-06-23', 'Sports',         5,  1150.00),
        ('2024-06-26', 'Electronics',    8,  5960.00),
        ('2024-06-28', 'Clothing',      16,  2496.00),
        # July 2024
        ('2024-07-02', 'Electronics',   19, 14155.00),
        ('2024-07-04', 'Clothing',      10,  1560.00),
        ('2024-07-06', 'Home & Garden',  7,  1190.00),
        ('2024-07-08', 'Books',         20,   599.80),
        ('2024-07-10', 'Sports',        25,  5750.00),
        ('2024-07-13', 'Electronics',   11,  8195.00),
        ('2024-07-15', 'Clothing',      14,  2184.00),
        ('2024-07-18', 'Home & Garden', 12,  2040.00),
        ('2024-07-20', 'Books',          8,   239.92),
        ('2024-07-23', 'Sports',         9,  2070.00),
        ('2024-07-26', 'Electronics',    5,  3725.00),
        ('2024-07-29', 'Clothing',       7,  1092.00),
        # August 2024
        ('2024-08-02', 'Electronics',   16, 11920.00),
        ('2024-08-04', 'Clothing',      12,  1872.00),
        ('2024-08-06', 'Home & Garden',  8,  1360.00),
        ('2024-08-08', 'Books',         22,   659.78),
        ('2024-08-10', 'Sports',        18,  4140.00),
        ('2024-08-13', 'Electronics',    9,  6705.00),
        ('2024-08-15', 'Clothing',       5,   780.00),
        ('2024-08-18', 'Home & Garden', 14,  2380.00),
        ('2024-08-20', 'Books',         30,   899.70),
        ('2024-08-23', 'Sports',        11,  2530.00),
        ('2024-08-26', 'Electronics',   21, 15645.00),
        ('2024-08-29', 'Clothing',       8,  1248.00),
        # September 2024
        ('2024-09-02', 'Electronics',   13,  9685.00),
        ('2024-09-04', 'Clothing',      16,  2496.00),
        ('2024-09-06', 'Home & Garden', 10,  1700.00),
        ('2024-09-08', 'Books',         18,   539.82),
        ('2024-09-10', 'Sports',         8,  1840.00),
        ('2024-09-13', 'Electronics',    7,  5215.00),
        ('2024-09-15', 'Clothing',       9,  1404.00),
        ('2024-09-18', 'Home & Garden',  6,  1020.00),
        ('2024-09-20', 'Books',         15,   449.85),
        ('2024-09-23', 'Sports',        13,  2990.00),
        ('2024-09-26', 'Electronics',   20, 14900.00),
        ('2024-09-29', 'Clothing',      11,  1716.00),
        # October 2024
        ('2024-10-02', 'Electronics',   25, 18625.00),
        ('2024-10-04', 'Clothing',      20,  3120.00),
        ('2024-10-06', 'Home & Garden', 15,  2550.00),
        ('2024-10-08', 'Books',         35,  1049.65),
        ('2024-10-10', 'Sports',        22,  5060.00),
        ('2024-10-13', 'Electronics',   14, 10430.00),
        ('2024-10-15', 'Clothing',      18,  2808.00),
        ('2024-10-18', 'Home & Garden', 10,  1700.00),
        ('2024-10-20', 'Books',         24,   719.76),
        ('2024-10-23', 'Sports',        16,  3680.00),
        ('2024-10-26', 'Electronics',   30, 22350.00),
        ('2024-10-29', 'Clothing',      12,  1872.00),
        # November 2024
        ('2024-11-01', 'Electronics',   40, 29800.00),
        ('2024-11-03', 'Clothing',      30,  4680.00),
        ('2024-11-05', 'Home & Garden', 20,  3400.00),
        ('2024-11-07', 'Books',         50,  1499.50),
        ('2024-11-09', 'Sports',        35,  8050.00),
        ('2024-11-11', 'Electronics',   55, 40975.00),
        ('2024-11-13', 'Clothing',      45,  7020.00),
        ('2024-11-15', 'Home & Garden', 25,  4250.00),
        ('2024-11-17', 'Books',         60,  1799.40),
        ('2024-11-19', 'Sports',        40,  9200.00),
        ('2024-11-22', 'Electronics',   35, 26075.00),
        ('2024-11-25', 'Clothing',      28,  4368.00),
        ('2024-11-28', 'Home & Garden', 18,  3060.00),
        # December 2024
        ('2024-12-02', 'Electronics',   30, 22350.00),
        ('2024-12-04', 'Clothing',      25,  3900.00),
        ('2024-12-06', 'Home & Garden', 16,  2720.00),
        ('2024-12-08', 'Books',         40,  1199.60),
        ('2024-12-10', 'Sports',        28,  6440.00),
        ('2024-12-13', 'Electronics',   20, 14900.00),
        ('2024-12-15', 'Clothing',      22,  3432.00),
        ('2024-12-18', 'Home & Garden', 12,  2040.00),
        ('2024-12-20', 'Books',         35,  1049.65),
        ('2024-12-23', 'Sports',        20,  4600.00),
        ('2024-12-26', 'Electronics',   45, 33525.00),
        ('2024-12-29', 'Clothing',      30,  4680.00),
    ]

    for r, row_data in enumerate(orders_data, 2):
        ws1.cell(row=r, column=1, value=row_data[0])
        ws1.cell(row=r, column=2, value=row_data[1])
        ws1.cell(row=r, column=3, value=row_data[2])
        ws1.cell(row=r, column=4, value=row_data[3])
        ws1.cell(row=r, column=4).number_format = '#,##0.00'

    # Set column widths for readability
    ws1.column_dimensions['A'].width = 14
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 12

    # Freeze header row
    ws1.freeze_panes = 'A2'

    # --- Sheet 2: Summary (empty - agent must fill this) ---
    ws2 = wb.create_sheet('Summary')
    # No data or formulas - agent must create the summary table

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
