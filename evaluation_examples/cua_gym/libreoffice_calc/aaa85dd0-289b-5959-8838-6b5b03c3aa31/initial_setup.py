"""
Initial Setup: Apply custom number format to display large market cap values with M suffix
Task ID: calc_fmt_numfmt_large_numbers_098
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_numfmt_large_numbers_098'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Stock Data ---
    ws = wb.active
    ws.title = 'Stock Data'

    # Headers
    ws['A1'] = 'Company'
    ws['B1'] = 'Market Cap'
    ws['C1'] = 'P/E Ratio'

    # Style headers with bold
    for col in ['A', 'B', 'C']:
        ws[f'{col}1'].font = Font(bold=True)

    # Stock data: realistic company names, market cap in absolute dollars (billions range),
    # and P/E ratios. Column B uses General format (no custom number format).
    # B2 = 4250000000, B3 = 890000000 as specified in context.
    data = [
        # Company,               Market Cap ($),   P/E Ratio
        ('Apple Inc.',           4250000000,        28.5),
        ('Nexify Technologies',  890000000,         15.2),
        ('DataStream Corp.',     1750000000,        22.8),
        ('GlobalRetail Holdings',3200000000,        19.4),
        ('HealthCore Systems',   620000000,         31.7),
        ('AeroPlex Industries',  2850000000,        17.3),
        ('ClearVision Media',    410000000,         24.1),
        ('PrimeTech Solutions',  1380000000,        20.6),
        ('Vertex Energy Ltd.',   5600000000,        14.9),
        ('NovaBio Sciences',     330000000,         45.3),
        ('Sterling Finance',     2100000000,        12.8),
        ('OmniCloud Networks',   780000000,         29.2),
        ('BlueStar Logistics',   1500000000,        18.5),
        ('IronGate Manufacturing',960000000,        16.7),
        ('Quantum Devices Inc.', 4800000000,        23.4),
        ('Pacific Consumer Goods',1120000000,       21.0),
        ('TerraFlex Agriculture', 490000000,        27.6),
        ('CityCore Real Estate',  2300000000,       11.3),
        ('Frontier Telecom Inc.', 3750000000,       16.1),
    ]

    for row_idx, (company, mktcap, pe) in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=company)
        ws.cell(row=row_idx, column=2, value=mktcap)
        ws.cell(row=row_idx, column=3, value=pe)
        # Column B uses General format — explicitly set to ensure no custom format
        ws.cell(row=row_idx, column=2).number_format = 'General'

    # Column widths for readability
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Stock Data')
    print(f'  Rows: 1 header + 19 data rows (rows 2-20)')
    print(f'  Column B: General format (large dollar values)')

create_initial()
