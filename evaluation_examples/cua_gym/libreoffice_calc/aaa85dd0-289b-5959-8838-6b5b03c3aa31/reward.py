"""
Reward Script: Apply custom number format to display large market cap values in millions with M suffix
Task ID: calc_fmt_numfmt_large_numbers_098
Domain: libreoffice_calc
Scoring:
  Component 1 (0.6): Cells B2:B20 ALL have custom number format that divides by 1M and appends "M"
                     (format must contain ",," to divide by 1,000,000 and end with M/\"M\" suffix)
                     This FAILS on initial (all 'General') and PASSES on golden ('#,##0,,"M"')
  Component 2 (0.4): Underlying numeric values in B2:B20 remain unchanged AND
                     no other cells (B1, columns A or C) were accidentally modified
                     This is a COMPOUND check: only passes when format is applied AND data integrity holds.
                     Scored only if Component 1 passes (cells have been formatted).
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — reward scripts run on the VM
TASK_ID = 'calc_fmt_numfmt_large_numbers_098'

# Expected values for B2:B20 (absolute dollar amounts, billions range)
EXPECTED_VALUES = {
    2: 4250000000,
    3: 890000000,
    4: 1750000000,
    5: 3200000000,
    6: 620000000,
    7: 2850000000,
    8: 410000000,
    9: 1380000000,
    10: 5600000000,
    11: 330000000,
    12: 2100000000,
    13: 780000000,
    14: 1500000000,
    15: 960000000,
    16: 4800000000,
    17: 1120000000,
    18: 490000000,
    19: 2300000000,
    20: 3750000000,
}


def is_millions_format(fmt_str):
    """
    Check if a number format string represents division by 1,000,000 with M suffix.
    Valid formats include: '#,##0,,"M"', '#,##0,,M', '0,,"M"', '0,,M', etc.
    The key indicators are:
      - ',,': two consecutive commas (scales by 10^6 in Excel/Calc number format spec)
      - 'M' or '"M"': the M suffix (millions)
    """
    if not fmt_str or fmt_str == 'General':
        return False
    # Must have double comma (divide by million) somewhere in format
    if ',,' not in fmt_str:
        return False
    # Must end with M (either quoted or unquoted)
    fmt_upper = fmt_str.upper()
    if not (fmt_upper.endswith('"M"') or fmt_upper.endswith('M')):
        return False
    return True


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: verify sheet exists
    if 'Stock Data' not in wb.sheetnames:
        print("FAIL: Sheet 'Stock Data' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Stock Data']

    # Component 1: ALL 19 cells B2:B20 have a valid millions custom format (0.6 points)
    # FAILS on initial file (all cells use 'General' format)
    # PASSES on golden file (all cells use '#,##0,,"M"')
    cells_with_format = []
    cells_without_format = []

    try:
        for row in range(2, 21):
            cell = ws.cell(row=row, column=2)
            fmt = cell.number_format
            if is_millions_format(fmt):
                cells_with_format.append(row)
            else:
                cells_without_format.append((row, fmt))

        if len(cells_with_format) == 19:
            # Sample format for reporting
            sample_fmt = ws.cell(row=2, column=2).number_format
            print(f"PASS: Component 1 — All 19 cells B2:B20 have millions custom format (e.g. {repr(sample_fmt)}) (0.6 pts)")
            total_score += 0.6
        elif len(cells_with_format) > 0:
            print(f"FAIL: Component 1 — Only {len(cells_with_format)}/19 cells have millions format. "
                  f"Missing cells: {['B'+str(r) for r,_ in cells_without_format[:5]]}")
        else:
            print(f"FAIL: Component 1 — No cells in B2:B20 have a valid millions custom format. "
                  f"Sample: B2={repr(ws.cell(row=2, column=2).number_format)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Underlying values unchanged AND no collateral modifications (0.4 points)
    # This is a COMPOUND check that only makes sense when formatting has been applied.
    # It verifies that while FORMAT changed, actual VALUES did not change.
    # FAILS on initial file when Component 1 also fails (no formatting applied at all)
    # PASSES on golden file: format applied + values are preserved.
    # NOTE: This only awards points when ALL 19 cells are formatted (cells_with_format == 19).
    # This ensures the 0.4 never fires alone on the initial file.
    try:
        if len(cells_with_format) == 19:
            # Check all values in B2:B20 remain as expected
            values_intact = True
            bad_value_info = None
            for row, expected_val in EXPECTED_VALUES.items():
                actual_val = ws.cell(row=row, column=2).value
                if actual_val != expected_val:
                    values_intact = False
                    bad_value_info = f"B{row}: expected {expected_val}, got {repr(actual_val)}"
                    break

            # Check that B1 header was not accidentally reformatted
            b1_fmt = ws.cell(row=1, column=2).number_format
            header_intact = (b1_fmt == 'General')

            if values_intact and header_intact:
                print(f"PASS: Component 2 — Underlying values preserved, B1 header format unchanged (0.4 pts)")
                total_score += 0.4
            elif not values_intact:
                print(f"FAIL: Component 2 — Underlying value was modified: {bad_value_info}")
            else:
                print(f"FAIL: Component 2 — B1 header format changed to {repr(b1_fmt)}, expected 'General'")
        else:
            print(f"SKIP: Component 2 — Skipped (Component 1 not fully passing; "
                  f"only {len(cells_with_format)}/19 cells formatted)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
