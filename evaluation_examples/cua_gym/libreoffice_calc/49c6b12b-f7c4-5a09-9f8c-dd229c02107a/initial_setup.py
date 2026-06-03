"""
Initial Setup: Create NPS survey spreadsheet with scores for conditional formatting task
Task ID: calc_gfl_072
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_072'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'NPS Scores'

    # --- Headers ---
    headers = ['Respondent', 'NPS Score', 'Category', 'Date', 'Channel']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FFD9E2F3', end_color='FFD9E2F3', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # --- Data: 29 survey responses (rows 2-30) ---
    respondents = [
        ('Sarah Chen', 9, 'Promoter', '2025-08-03', 'Email'),
        ('Marcus Johnson', 7, 'Passive', '2025-08-04', 'In-App'),
        ('Priya Patel', 10, 'Promoter', '2025-08-04', 'Email'),
        ('James O\'Brien', 3, 'Detractor', '2025-08-05', 'Phone'),
        ('Aisha Mohammed', 8, 'Passive', '2025-08-05', 'Email'),
        ('Carlos Rivera', 6, 'Detractor', '2025-08-06', 'In-App'),
        ('Emily Watson', 10, 'Promoter', '2025-08-07', 'Email'),
        ('Dmitri Volkov', 2, 'Detractor', '2025-08-07', 'Phone'),
        ('Lisa Nakamura', 9, 'Promoter', '2025-08-08', 'In-App'),
        ('Thomas Mueller', 5, 'Detractor', '2025-08-08', 'Email'),
        ('Fatima Al-Rashid', 8, 'Passive', '2025-08-09', 'Phone'),
        ('Daniel Kim', 1, 'Detractor', '2025-08-09', 'In-App'),
        ('Rachel Green', 9, 'Promoter', '2025-08-10', 'Email'),
        ('Antonio Rossi', 7, 'Passive', '2025-08-10', 'Phone'),
        ('Mei-Ling Wu', 10, 'Promoter', '2025-08-11', 'Email'),
        ('Robert Taylor', 4, 'Detractor', '2025-08-11', 'In-App'),
        ('Sofia Hernandez', 8, 'Passive', '2025-08-12', 'Email'),
        ('Henrik Larsson', 0, 'Detractor', '2025-08-12', 'Phone'),
        ('Yuki Tanaka', 9, 'Promoter', '2025-08-13', 'In-App'),
        ('Grace Okafor', 6, 'Detractor', '2025-08-13', 'Email'),
        ('Alexandre Dubois', 10, 'Promoter', '2025-08-14', 'Phone'),
        ('Nadia Petrov', 3, 'Detractor', '2025-08-14', 'In-App'),
        ('Michael Chang', 8, 'Passive', '2025-08-15', 'Email'),
        ('Isabella Costa', 7, 'Passive', '2025-08-15', 'Phone'),
        ('Omar Hassan', 5, 'Detractor', '2025-08-16', 'In-App'),
        ('Lena Johansson', 9, 'Promoter', '2025-08-16', 'Email'),
        ('David Nguyen', 2, 'Detractor', '2025-08-17', 'Phone'),
        ('Amara Williams', 10, 'Promoter', '2025-08-17', 'Email'),
        ('Kenji Watanabe', 6, 'Detractor', '2025-08-18', 'In-App'),
    ]

    for r, (name, score, category, date, channel) in enumerate(respondents, 2):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=score)
        ws.cell(row=r, column=3, value=category)
        ws.cell(row=r, column=4, value=date)
        ws.cell(row=r, column=5, value=channel)

    # --- Column widths ---
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12

    # --- NO conditional formatting in initial state ---

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
