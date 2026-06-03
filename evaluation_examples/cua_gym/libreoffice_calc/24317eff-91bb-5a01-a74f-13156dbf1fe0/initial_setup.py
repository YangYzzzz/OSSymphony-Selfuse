"""
Initial Setup: Apply orange background to top 5 values in D2:D40
Task ID: calc_gfl_079
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_079'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scores"

    # --- Headers ---
    headers = ["Participant", "Category", "Round", "Score", "Rank"]
    header_font = Font(name="Calibri", size=11, bold=True)
    thin = Side(style="thin", color="000000")
    header_border = Border(bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10

    # --- Data (39 rows: rows 2-40) ---
    participants = [
        "Sarah Chen", "Marcus Johnson", "Elena Rodriguez", "James O'Brien",
        "Aiko Tanaka", "David Kim", "Priya Patel", "Lucas Fernandez",
        "Olivia Williams", "Noah Garcia", "Sophia Martinez", "Liam Anderson",
        "Isabella Thomas", "Ethan Jackson", "Mia White", "Alexander Harris",
        "Charlotte Martin", "Benjamin Thompson", "Amelia Robinson", "William Clark",
        "Harper Lewis", "Daniel Walker", "Evelyn Hall", "Michael Allen",
        "Abigail Young", "Henry Hernandez", "Emily King", "Sebastian Wright",
        "Victoria Lopez", "Owen Hill", "Grace Scott", "Jack Green",
        "Chloe Adams", "Ryan Baker", "Lily Nelson", "Nathan Carter",
        "Zoey Mitchell", "Samuel Perez", "Aria Roberts"
    ]

    categories = [
        "Sprint", "Endurance", "Agility", "Precision", "Strength",
        "Balance", "Speed", "Flexibility"
    ]

    # Scores designed so the top 5 are clearly identifiable: 99, 97, 96, 95, 94
    # The rest range from 45 to 93
    scores = [
        78, 65, 82, 99, 71, 58, 88, 45, 93, 67,
        96, 54, 76, 62, 85, 49, 91, 73, 60, 87,
        52, 97, 69, 83, 56, 79, 64, 95, 47, 90,
        74, 68, 81, 57, 94, 63, 86, 51, 72
    ]

    rounds_data = [
        1, 2, 3, 1, 2, 3, 1, 2, 3, 1,
        2, 3, 1, 2, 3, 1, 2, 3, 1, 2,
        3, 1, 2, 3, 1, 2, 3, 1, 2, 3,
        1, 2, 3, 1, 2, 3, 1, 2, 3
    ]

    ranks = [
        12, 22, 8, 1, 17, 28, 5, 39, 4, 20,
        3, 31, 14, 24, 7, 36, 6, 16, 26, 6,
        33, 2, 19, 9, 30, 11, 23, 3, 37, 7,
        15, 21, 10, 29, 4, 25, 6, 34, 18
    ]

    for i in range(39):
        row = i + 2
        ws.cell(row=row, column=1, value=participants[i])
        ws.cell(row=row, column=2, value=categories[i % len(categories)])
        ws.cell(row=row, column=3, value=rounds_data[i])
        ws.cell(row=row, column=4, value=scores[i])
        ws.cell(row=row, column=5, value=ranks[i])

    # NO conditional formatting in initial state

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
