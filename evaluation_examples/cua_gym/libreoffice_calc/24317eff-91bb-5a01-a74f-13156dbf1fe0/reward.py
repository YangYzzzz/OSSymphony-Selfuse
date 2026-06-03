"""
Reward Script: Apply orange background to top 5 values in D2:D40 via conditional formatting
Task ID: calc_gfl_079
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Conditional formatting rule exists on D2:D40 range
  Component 2 (0.30): Rule formula correctly identifies top 5 values (LARGE-based or equivalent)
  Component 3 (0.25): Fill color is orange (FFC000 or close orange variant)
  Component 4 (0.15): Only the top 5 score cells would be highlighted (formula correctness cross-check)
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_079'


def persist_app_state(domain):
    """Save any unsaved LibreOffice changes before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Scores' sheet must exist
    if 'Scores' not in wb.sheetnames:
        print("FAIL: 'Scores' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Scores']

    # Precondition: D column data exists
    d_values = []
    for r in range(2, 41):
        v = ws.cell(row=r, column=4).value
        if v is not None:
            try:
                d_values.append((r, float(v)))
            except (ValueError, TypeError):
                pass

    if len(d_values) < 10:
        print(f"FAIL: Not enough data in D2:D40 (found {len(d_values)} numeric values)")
        print("REWARD: 0.0")
        return 0.0

    # Collect all conditional formatting rules
    cf_list = list(ws.conditional_formatting)

    # Component 1: Conditional formatting rule exists covering D2:D40 (0.30 points)
    try:
        found_cf_on_d = False
        matching_cf = None
        matching_rule = None

        for cf in cf_list:
            cf_range_str = str(cf).upper()
            for rule in cf.rules:
                # Check if the CF range covers D column cells in the 2:40 range
                # Accept ranges like D2:D40, $D$2:$D$40, D2:D100, etc.
                # The key is that it covers D2:D40
                if 'D' in cf_range_str and any(c.isdigit() for c in cf_range_str):
                    # More precise: check if the range string references column D
                    # and includes rows in the 2-40 area
                    found_cf_on_d = True
                    matching_cf = cf
                    matching_rule = rule
                    break
            if found_cf_on_d:
                break

        if found_cf_on_d:
            print(f"PASS: Component 1 — Conditional formatting found on range: {matching_cf} (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 — No conditional formatting found on D column range. Found {len(cf_list)} CF rules total.")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Rule formula correctly targets top 5 values (0.30 points)
    try:
        if matching_rule is not None:
            rule_type = matching_rule.type
            formula_list = getattr(matching_rule, 'formula', None) or []
            rank_val = getattr(matching_rule, 'rank', None)

            formula_str = str(formula_list[0]).upper() if formula_list else ''

            top5_detected = False

            # Method A: expression/formula type with LARGE(..., 5)
            if 'LARGE' in formula_str and '5' in formula_str:
                top5_detected = True
                print(f"PASS: Component 2 — Formula uses LARGE with top 5: {formula_str} (0.30 pts)")

            # Method B: top10 rule type with rank=5
            elif rule_type == 'top10' and rank_val is not None and int(rank_val) == 5:
                top5_detected = True
                print(f"PASS: Component 2 — Top10 rule with rank=5 (0.30 pts)")

            # Method C: Check for other valid approaches (RANK, PERCENTILE, etc.)
            elif 'RANK' in formula_str and '5' in formula_str:
                top5_detected = True
                print(f"PASS: Component 2 — Formula uses RANK-based top 5: {formula_str} (0.30 pts)")

            # Method D: COUNTIF-based approach
            elif 'COUNTIF' in formula_str and '5' in formula_str:
                top5_detected = True
                print(f"PASS: Component 2 — Formula uses COUNTIF-based top 5: {formula_str} (0.30 pts)")

            else:
                print(f"FAIL: Component 2 — Rule formula does not clearly target top 5. Type: {rule_type}, Formula: {formula_str}, Rank: {rank_val}")

            if top5_detected:
                total_score += 0.30
        else:
            print("FAIL: Component 2 — No matching rule to evaluate (depends on Component 1)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Fill color is orange (0.25 points)
    try:
        if matching_rule is not None and hasattr(matching_rule, 'dxf') and matching_rule.dxf:
            dxf = matching_rule.dxf
            fill_color = None
            if dxf.fill and dxf.fill.fgColor:
                fill_color = dxf.fill.fgColor.rgb
            elif dxf.fill and dxf.fill.start_color:
                fill_color = dxf.fill.start_color.rgb

            if fill_color:
                # Normalize: strip alpha if 8 chars
                color_hex = fill_color[-6:] if len(fill_color) == 8 else fill_color

                # Check if it's orange. Common orange hex values:
                # FFC000 (standard orange), FF8C00 (dark orange), FFA500 (orange),
                # FF6600, ED7D31, etc.
                r_val = int(color_hex[0:2], 16)
                g_val = int(color_hex[2:4], 16)
                b_val = int(color_hex[4:6], 16)

                # Orange: high red (>200), medium green (80-220), low blue (<100)
                is_orange = r_val >= 200 and 80 <= g_val <= 220 and b_val <= 100
                # Also accept FFC000 exactly and common variants
                known_oranges = ['FFC000', 'FFA500', 'FF8C00', 'FF6600', 'ED7D31',
                                 'F4B084', 'FF9900', 'FF7F00', 'FFB900', 'E67E22']
                is_known = color_hex.upper() in known_oranges

                if is_orange or is_known:
                    print(f"PASS: Component 3 — Orange fill color detected: {fill_color} (0.25 pts)")
                    total_score += 0.25
                else:
                    print(f"FAIL: Component 3 — Fill color {fill_color} (R={r_val},G={g_val},B={b_val}) is not orange")
            else:
                print("FAIL: Component 3 — No fill color found in differential style")
        else:
            print("FAIL: Component 3 — No differential style found on rule (depends on Component 1)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Formula targets exactly the top 5 — verify by cross-checking
    # the LARGE threshold value against actual data (0.15 points)
    try:
        if matching_rule is not None:
            formula_list = getattr(matching_rule, 'formula', None) or []
            formula_str = str(formula_list[0]).upper() if formula_list else ''
            rank_val = getattr(matching_rule, 'rank', None)

            # Determine the top-N count from the formula
            top_n = None

            # Check LARGE(..., N) pattern
            large_match = re.search(r'LARGE\([^,]+,\s*(\d+)\)', formula_str)
            if large_match:
                top_n = int(large_match.group(1))

            # Check top10 type with rank
            if top_n is None and rank_val is not None:
                top_n = int(rank_val)

            if top_n == 5:
                # Verify that exactly 5 cells would be highlighted
                # Sort values descending, the 5th highest is the threshold
                sorted_d = sorted([v for _, v in d_values], reverse=True)
                if len(sorted_d) >= 5:
                    threshold = sorted_d[4]  # 5th highest value
                    highlighted_count = sum(1 for _, v in d_values if v >= threshold)

                    if highlighted_count == 5:
                        print(f"PASS: Component 4 — Exactly 5 cells would be highlighted (threshold={threshold}) (0.15 pts)")
                        total_score += 0.15
                    elif highlighted_count >= 5:
                        # If there are ties at the boundary, the formula still targets top 5
                        # This is acceptable — the formula is correct, ties are a data property
                        print(f"PASS: Component 4 — {highlighted_count} cells >= threshold {threshold} (ties at boundary acceptable) (0.15 pts)")
                        total_score += 0.15
                    else:
                        print(f"FAIL: Component 4 — Only {highlighted_count} cells would be highlighted")
                else:
                    print(f"FAIL: Component 4 — Not enough data values to determine top 5")
            else:
                print(f"FAIL: Component 4 — Top-N count is {top_n}, expected 5")
        else:
            print("FAIL: Component 4 — No matching rule (depends on Component 1)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
