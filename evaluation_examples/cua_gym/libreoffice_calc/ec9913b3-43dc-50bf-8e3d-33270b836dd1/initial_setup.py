"""
Initial Setup: Research Project Budget Allocation
Task ID: calc_edu_research_budget_allocation_037
Domain: libreoffice_calc

Creates a ResearchBudget sheet with categories (Personnel, Equipment, Supplies,
Travel, Publication, Overhead) with Allocated and Committed columns populated,
but Remaining (D) and Pct of Total (E) columns empty (to be filled by agent).
No formulas in row 8 totals. No sheet protection applied.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_research_budget_allocation_037'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ResearchBudget'

    # --- Header row ---
    headers = ['Category', 'Allocated', 'Committed', 'Remaining', 'Pct of Total']
    header_font = Font(name='Calibri', bold=True, size=11)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', bold=True, size=11, color='FFFFFFFF')
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # --- Data rows: Category, Allocated, Committed (D and E intentionally empty) ---
    # Research project: NSF-funded computational neuroscience study
    data = [
        ('Personnel',    185000, 142300),
        ('Equipment',     62500,  58900),
        ('Supplies',      18750,  11430),
        ('Travel',        12000,   7650),
        ('Publication',    5500,   2200),
        ('Overhead',      46250,  35640),
    ]

    data_font = Font(name='Calibri', size=11)
    money_format = '#,##0.00'
    center_align = Alignment(horizontal='center')
    right_align = Alignment(horizontal='right')

    for r, (cat, allocated, committed) in enumerate(data, 2):
        # Column A: Category
        cell_a = ws.cell(row=r, column=1, value=cat)
        cell_a.font = data_font
        cell_a.border = border

        # Column B: Allocated
        cell_b = ws.cell(row=r, column=2, value=allocated)
        cell_b.font = data_font
        cell_b.number_format = money_format
        cell_b.alignment = right_align
        cell_b.border = border

        # Column C: Committed
        cell_c = ws.cell(row=r, column=3, value=committed)
        cell_c.font = data_font
        cell_c.number_format = money_format
        cell_c.alignment = right_align
        cell_c.border = border

        # Column D: Remaining — intentionally EMPTY (agent must add =B-C formula)
        cell_d = ws.cell(row=r, column=4, value=None)
        cell_d.font = data_font
        cell_d.number_format = money_format
        cell_d.border = border

        # Column E: Pct of Total — intentionally EMPTY (agent must add =B/$B$8 formula)
        cell_e = ws.cell(row=r, column=5, value=None)
        cell_e.font = data_font
        cell_e.number_format = '0.00%'
        cell_e.border = border

    # --- Row 8: Totals row (label in A, no formulas) ---
    totals_font = Font(name='Calibri', bold=True, size=11)
    totals_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')

    cell_a8 = ws.cell(row=8, column=1, value='Total')
    cell_a8.font = totals_font
    cell_a8.fill = totals_fill
    cell_a8.border = border

    # B8 and C8: no formulas yet (agent must add SUM)
    for col in [2, 3, 4, 5]:
        cell = ws.cell(row=8, column=col, value=None)
        cell.font = totals_font
        cell.fill = totals_fill
        cell.border = border
        if col in [2, 3, 4]:
            cell.number_format = money_format
        elif col == 5:
            cell.number_format = '0.00%'

    # --- Column widths ---
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14

    # --- Row height for header ---
    ws.row_dimensions[1].height = 20

    # No protection on the initial file

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheet: ResearchBudget')
    print('Rows 2-7: 6 budget categories with Allocated and Committed amounts')
    print('Columns D and E: empty (to be filled by agent)')
    print('Row 8: Totals label, no formulas')
    print('No sheet protection applied')


create_initial()
