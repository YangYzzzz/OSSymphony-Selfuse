"""
Reward Script: Allocate research project budget with formulas, cell locking, and sheet protection.
Task ID: calc_edu_research_budget_allocation_037
Domain: libreoffice_calc
Scoring:
  Component 1: Column D (Remaining) formulas =B-C for rows 2-7        (0.25 pts)
  Component 2: Column E (Pct of Total) formulas =B/$B$8 for rows 2-7  (0.25 pts)
  Component 3: Row 8 SUM formulas for B8, C8, D8, E8                  (0.20 pts)
  Component 4: C2:C7 unlocked; A2:A8, B2:B8, D2:D8, E2:E8 locked     (0.15 pts)
  Component 5: Sheet protection enabled                                 (0.15 pts)
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_research_budget_allocation_037'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'ResearchBudget' not in wb.sheetnames:
        print("CRITICAL: Sheet 'ResearchBudget' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['ResearchBudget']

    # Component 1: Column D (Remaining) has =B-C formulas for rows 2-7 (0.25 pts)
    # This fails on initial file (D2:D7 are all None) and passes on golden (formulas present)
    try:
        d_formula_count = 0
        for row in range(2, 8):
            val = ws.cell(row=row, column=4).value  # Column D
            if val is not None and isinstance(val, str):
                # Accept =Bx-Cx pattern (case-insensitive, with or without spaces)
                formula_clean = val.upper().replace(' ', '')
                expected = f'=B{row}-C{row}'
                if formula_clean == expected.upper():
                    d_formula_count += 1
                else:
                    print(f"FAIL: Component 1 — D{row} expected '{expected}', found: {repr(val)}")
            else:
                print(f"FAIL: Component 1 — D{row} expected formula, found: {repr(val)}")
        if d_formula_count == 6:
            print(f"PASS: Component 1 — All 6 Remaining formulas present in D2:D7 (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — Only {d_formula_count}/6 Remaining formulas found in D2:D7")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Column E (Pct of Total) has =B/$B$8 formulas for rows 2-7 (0.25 pts)
    # This fails on initial file (E2:E7 are all None) and passes on golden (formulas present)
    try:
        e_formula_count = 0
        for row in range(2, 8):
            val = ws.cell(row=row, column=5).value  # Column E
            if val is not None and isinstance(val, str):
                formula_clean = val.upper().replace(' ', '')
                expected = f'=B{row}/$B$8'
                if formula_clean == expected.upper():
                    e_formula_count += 1
                else:
                    print(f"FAIL: Component 2 — E{row} expected '{expected}', found: {repr(val)}")
            else:
                print(f"FAIL: Component 2 — E{row} expected formula, found: {repr(val)}")
        if e_formula_count == 6:
            print(f"PASS: Component 2 — All 6 Pct of Total formulas present in E2:E7 (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — Only {e_formula_count}/6 Pct of Total formulas found in E2:E7")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Row 8 has SUM formulas for B8, C8, D8, E8 (0.20 pts)
    # This fails on initial file (B8, C8, D8, E8 are all None) and passes on golden
    try:
        sum_cells = {
            'B8': ws.cell(row=8, column=2).value,
            'C8': ws.cell(row=8, column=3).value,
            'D8': ws.cell(row=8, column=4).value,
            'E8': ws.cell(row=8, column=5).value,
        }
        expected_sums = {
            'B8': '=SUM(B2:B7)',
            'C8': '=SUM(C2:C7)',
            'D8': '=SUM(D2:D7)',
            'E8': '=SUM(E2:E7)',
        }
        sum_count = 0
        for coord, val in sum_cells.items():
            expected = expected_sums[coord]
            if val is not None and isinstance(val, str) and val.upper().replace(' ', '') == expected.upper().replace(' ', ''):
                sum_count += 1
            else:
                print(f"FAIL: Component 3 — {coord} expected '{expected}', found: {repr(val)}")
        if sum_count == 4:
            print(f"PASS: Component 3 — All 4 SUM formulas present in B8:E8 (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 3 — Only {sum_count}/4 SUM formulas found in row 8")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: C2:C7 unlocked; A2:A8, B2:B8, D2:D8, E2:E8 locked (0.15 pts)
    # Initial file: all cells are locked=True (default). Golden: C2:C7 set to locked=False.
    try:
        # Check C2:C7 are unlocked (locked=False)
        c_unlocked_count = 0
        for row in range(2, 8):
            cell = ws.cell(row=row, column=3)  # Column C
            if cell.protection.locked == False:
                c_unlocked_count += 1
            else:
                print(f"FAIL: Component 4 — C{row} should be unlocked but locked={cell.protection.locked}")

        # Check A, B, D, E columns in rows 2-8 are locked
        locked_count = 0
        locked_total = 0
        cols_to_check = [1, 2, 4, 5]  # A, B, D, E
        for row in range(2, 9):
            for col in cols_to_check:
                locked_total += 1
                cell = ws.cell(row=row, column=col)
                if cell.protection.locked != False:  # locked=True or None (default is locked)
                    locked_count += 1
                else:
                    from openpyxl.utils import get_column_letter
                    coord = f"{get_column_letter(col)}{row}"
                    print(f"FAIL: Component 4 — {coord} should be locked but locked={cell.protection.locked}")

        if c_unlocked_count == 6 and locked_count == locked_total:
            print(f"PASS: Component 4 — C2:C7 unlocked ({c_unlocked_count}/6), A/B/D/E rows 2-8 locked ({locked_count}/{locked_total}) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — C2:C7 unlocked: {c_unlocked_count}/6, A/B/D/E locked: {locked_count}/{locked_total}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Sheet protection is enabled (0.15 pts)
    # Initial file: ws.protection.sheet = False. Golden: ws.protection.sheet = True.
    try:
        protection_enabled = ws.protection.sheet
        if protection_enabled:
            print(f"PASS: Component 5 — Sheet protection is enabled (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 5 — Sheet protection is not enabled (ws.protection.sheet={protection_enabled})")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
