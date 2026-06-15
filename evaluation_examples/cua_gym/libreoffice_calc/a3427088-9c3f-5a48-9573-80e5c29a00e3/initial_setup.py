"""
Initial Setup: Journal Submission Tracker - Faculty Research Submissions
Task ID: calc_edu_journal_submission_tracker_065
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_journal_submission_tracker_065'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Submissions ---
    ws = wb.active
    ws.title = 'Submissions'

    # Headers in row 1
    headers = [
        'Faculty Name',   # A
        'Journal',        # B
        'Submission Date', # C
        'Decision Date',  # D
        'Status',         # E
        'Days Under Review', # F - EMPTY, to be filled by agent
        'Long Wait Flag', # G - EMPTY, to be filled by agent
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Reference date in J1 (used in the formula)
    ws['J1'] = date(2025, 12, 15)
    ws['J1'].number_format = 'YYYY-MM-DD'

    # Realistic faculty submission data
    # Status options: 'Under Review', 'Accepted', 'Rejected', 'Revise & Resubmit'
    # F and G are intentionally blank - agent must fill these
    submissions = [
        # Faculty Name, Journal, Submission Date, Decision Date, Status
        ('Dr. Sarah Chen',         'Journal of Educational Psychology',          date(2025, 3, 10),  date(2025, 7, 22),  'Accepted'),
        ('Prof. Marcus Johnson',   'American Educational Research Journal',      date(2025, 1, 15),  date(2025, 5, 30),  'Rejected'),
        ('Dr. Priya Patel',        'Review of Educational Research',             date(2025, 7, 5),   None,               'Under Review'),
        ('Dr. James Okafor',       'Journal of Learning Sciences',               date(2025, 2, 20),  date(2025, 6, 14),  'Revise & Resubmit'),
        ('Prof. Elena Vasquez',    'Educational Researcher',                     date(2025, 8, 12),  None,               'Under Review'),
        ('Dr. Thomas Kim',         'British Journal of Educational Technology',  date(2025, 4, 3),   date(2025, 9, 18),  'Accepted'),
        ('Prof. Natasha Brennan',  'Journal of Curriculum Studies',              date(2025, 6, 22),  None,               'Under Review'),
        ('Dr. William Nguyen',     'Cognition and Instruction',                  date(2025, 9, 14),  None,               'Under Review'),
        ('Prof. Diana Osei',       'Teaching and Teacher Education',             date(2025, 3, 28),  date(2025, 8, 5),   'Accepted'),
        ('Dr. Carlos Rivera',      'Journal of Educational Research',            date(2025, 5, 17),  date(2025, 10, 2),  'Revise & Resubmit'),
        ('Prof. Linda Hoffman',    'Learning and Instruction',                   date(2025, 7, 30),  None,               'Under Review'),
        ('Dr. Ahmed Hassan',       'Educational Studies in Mathematics',         date(2025, 2, 8),   date(2025, 5, 19),  'Rejected'),
        ('Prof. Rebecca Stone',    'Journal of Research in Science Teaching',    date(2025, 8, 25),  None,               'Under Review'),
        ('Dr. Kevin Park',         'Contemporary Educational Psychology',        date(2025, 1, 30),  date(2025, 6, 11),  'Accepted'),
        ('Prof. Maria Santos',     'International Journal of Science Education', date(2025, 6, 5),   None,               'Under Review'),
        ('Dr. Jonathan Blake',     'Journal of Educational Psychology',          date(2025, 4, 18),  date(2025, 9, 27),  'Rejected'),
        ('Prof. Fatima Al-Rashid', 'American Educational Research Journal',      date(2025, 9, 3),   None,               'Under Review'),
        ('Dr. Samuel Wright',      'Educational Psychologist',                   date(2025, 3, 22),  date(2025, 7, 8),   'Accepted'),
        ('Prof. Grace Liu',        'Journal of Learning Sciences',               date(2025, 8, 8),   None,               'Under Review'),
        ('Dr. Patrick Morley',     'Review of Educational Research',             date(2025, 5, 2),   date(2025, 9, 15),  'Revise & Resubmit'),
        ('Prof. Ananya Sharma',    'Instructional Science',                      date(2025, 7, 18),  None,               'Under Review'),
        ('Dr. Michael Torres',     'Journal of Educational Research',            date(2025, 2, 14),  date(2025, 6, 29),  'Accepted'),
        ('Prof. Julia Erikson',    'Educational Researcher',                     date(2025, 9, 20),  None,               'Under Review'),
        ('Dr. Benjamin Adeyemi',   'Computers & Education',                      date(2025, 4, 7),   date(2025, 8, 22),  'Accepted'),
        ('Prof. Isabella Crawford', 'Teaching and Teacher Education',            date(2025, 6, 15),  None,               'Under Review'),
        ('Dr. Andrew Tan',         'Learning and Instruction',                   date(2025, 1, 25),  date(2025, 5, 14),  'Rejected'),
        ('Prof. Sandra Williams',  'Journal of Curriculum Studies',              date(2025, 8, 1),   None,               'Under Review'),
        ('Dr. Hiroshi Yamamoto',   'British Journal of Educational Technology',  date(2025, 3, 5),   date(2025, 7, 31),  'Accepted'),
        ('Prof. Claire Dubois',    'Cognition and Instruction',                  date(2025, 7, 10),  None,               'Under Review'),
        ('Dr. Richard Okonkwo',    'Educational Studies in Mathematics',         date(2025, 5, 28),  date(2025, 10, 10), 'Revise & Resubmit'),
        ('Prof. Mei-Ling Zhou',    'International Journal of Science Education', date(2025, 9, 8),   None,               'Under Review'),
        ('Dr. Laura Beckett',      'Contemporary Educational Psychology',        date(2025, 2, 3),   date(2025, 5, 27),  'Accepted'),
        ('Prof. Dmitri Sokolov',   'Journal of Research in Science Teaching',    date(2025, 8, 18),  None,               'Under Review'),
        ('Dr. Nicole Lambert',     'Educational Psychologist',                   date(2025, 4, 25),  date(2025, 9, 5),   'Revise & Resubmit'),
        ('Prof. Omar Abdullah',    'Journal of Educational Psychology',          date(2025, 6, 30),  None,               'Under Review'),
        ('Dr. Patricia Walsh',     'American Educational Research Journal',      date(2025, 1, 10),  date(2025, 4, 30),  'Rejected'),
        ('Prof. Liang Chen',       'Instructional Science',                      date(2025, 7, 22),  None,               'Under Review'),
        ('Dr. Stefan Mueller',     'Computers & Education',                      date(2025, 3, 15),  date(2025, 8, 12),  'Accepted'),
        ('Prof. Amelia Nguyen',    'Journal of Learning Sciences',               date(2025, 9, 28),  None,               'Under Review'),
        ('Dr. Charles Osei',       'Review of Educational Research',             date(2025, 5, 8),   date(2025, 9, 22),  'Accepted'),
        ('Prof. Roshni Kapoor',    'Educational Researcher',                     date(2025, 8, 5),   None,               'Under Review'),
        ('Dr. Nathan Brooks',      'Journal of Curriculum Studies',              date(2025, 2, 28),  date(2025, 7, 15),  'Rejected'),
        ('Prof. Sylvia Ferreira',  'Teaching and Teacher Education',             date(2025, 6, 12),  None,               'Under Review'),
        ('Dr. Theodore Chu',       'Learning and Instruction',                   date(2025, 9, 15),  None,               'Under Review'),
        ('Prof. Ingrid Hansen',    'British Journal of Educational Technology',  date(2025, 4, 20),  date(2025, 9, 3),   'Accepted'),
    ]

    for r, (faculty, journal, sub_date, dec_date, status) in enumerate(submissions, 2):
        ws.cell(row=r, column=1, value=faculty)
        ws.cell(row=r, column=2, value=journal)

        c_cell = ws.cell(row=r, column=3, value=sub_date)
        c_cell.number_format = 'YYYY-MM-DD'

        d_cell = ws.cell(row=r, column=4, value=dec_date)
        if dec_date is not None:
            d_cell.number_format = 'YYYY-MM-DD'

        ws.cell(row=r, column=5, value=status)
        # Columns F and G are intentionally left empty (Days Under Review, Long Wait Flag)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Submissions')
    print(f'  Rows 2-46: 45 submissions (F and G columns empty)')
    print(f'  J1: reference date 2025-12-15')


create_initial()
