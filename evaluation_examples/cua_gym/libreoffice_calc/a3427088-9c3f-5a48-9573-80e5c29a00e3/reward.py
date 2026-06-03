"""
Reward Script: Track faculty research journal submissions
Task ID: calc_edu_journal_submission_tracker_065
Domain: libreoffice_calc
Scoring:
  - Component 1: Days Under Review formulas (F2:F46) — 0.30 pts
  - Component 2: Long Wait Flag formulas (G2:G46) — 0.20 pts
  - Component 3: Status summary table (I3:J8) with COUNTIF — 0.20 pts
  - Component 4: Faculty summary table (I11:J57) with COUNTIF — 0.20 pts
  - Component 5: Conditional formatting yellow fill for Follow Up rows — 0.10 pts
  Total: 1.00
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_journal_submission_tracker_065'


def count_days_formulas(ws):
    """Count rows 2-46 in column F with =IF(D?="",J$1-C?,D?-C?) pattern."""
    found = 0
    for row in range(2, 47):
        val = ws.cell(row=row, column=6).value  # Column F
        if isinstance(val, str) and val.startswith('=IF(') and 'J$1' in val and '-C' in val:
            found += 1
    return found


def count_flag_formulas(ws):
    """Count rows 2-46 in column G with =IF(AND(E?="Under Review",F?>90),"Follow Up","") pattern."""
    found = 0
    for row in range(2, 47):
        val = ws.cell(row=row, column=7).value  # Column G
        if (isinstance(val, str) and val.startswith('=IF(AND(')
                and 'Under Review' in val and '>90' in val and 'Follow Up' in val):
            found += 1
    return found


def count_status_countif(ws):
    """Count COUNTIF formulas in J5:J8 referencing $E$ column (status summary)."""
    status_values = ['Under Review', 'Accepted', 'Rejected', 'Revise & Resubmit']
    found = 0
    for row_idx, expected_status in zip(range(5, 9), status_values):
        i_val = ws.cell(row=row_idx, column=9).value
        j_val = ws.cell(row=row_idx, column=10).value
        if (i_val == expected_status
                and isinstance(j_val, str)
                and 'COUNTIF' in j_val.upper()
                and '$E$' in j_val):
            found += 1
    return found


def count_faculty_countif(ws):
    """Count COUNTIF formulas in J13:J57 referencing $A$ column (faculty summary)."""
    found = 0
    for row_idx in range(13, 58):
        j_val = ws.cell(row_idx, 10).value
        if isinstance(j_val, str) and 'COUNTIF' in j_val.upper() and '$A$' in j_val:
            found += 1
    return found


def score_conditional_formatting(ws):
    """
    Check if there is a conditional formatting rule that references
    column G for 'Follow Up' and applies a fill color.
    Returns: 0.10 if rule + fill color found, 0.05 if rule found without color, 0.0 otherwise.
    """
    cf_rules = ws.conditional_formatting
    rules_with_color = 0
    rules_without_color = 0
    for cf_range in cf_rules:
        for rule in cf_range.rules:
            if rule.type != 'expression':
                continue
            if not rule.formula or len(rule.formula) == 0:
                continue
            formula_str = str(rule.formula[0]).replace(' ', '')
            if 'G' not in formula_str or 'Follow' not in formula_str:
                continue
            # Rule references G and Follow Up — now check fill
            dxf = rule.dxf
            has_fill_color = 0
            if dxf and dxf.fill and dxf.fill.fgColor:
                try:
                    fill_color = dxf.fill.fgColor.rgb
                    if fill_color and len(fill_color) >= 6:
                        has_fill_color = 1
                except Exception:
                    pass
            if has_fill_color == 1:
                rules_with_color += 1
            else:
                rules_without_color += 1

    if rules_with_color >= 1:
        return 0.10
    if rules_without_color >= 1:
        return 0.05
    return 0.0


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Submissions' not in wb.sheetnames:
        print("CRITICAL: 'Submissions' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Submissions']

    # Component 1: Days Under Review formulas in F2:F46 (0.30 points)
    # Pattern: =IF(D?="",J$1-C?,D?-C?)
    # FAILS on initial (F column is empty) → PASSES on golden (F2:F46 have IF formulas)
    try:
        f_count = count_days_formulas(ws)
        f_total = 45  # rows 2-46
        if f_count == f_total:
            print(f"PASS: Component 1 — All {f_count}/{f_total} Days Under Review formulas present (0.30 pts)")
            total_score += 0.30
        elif f_count >= 1:
            c1_partial = round(0.30 * f_count / f_total, 4)
            print(f"PARTIAL: Component 1 — {f_count}/{f_total} Days Under Review formulas present ({c1_partial} pts)")
            total_score += c1_partial
        else:
            print(f"FAIL: Component 1 — Expected IF(D?=\"\",J$1-C?,D?-C?) formulas in F2:F46, found {f_count}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Long Wait Flag formulas in G2:G46 (0.20 points)
    # Pattern: =IF(AND(E?="Under Review",F?>90),"Follow Up","")
    # FAILS on initial (G column is empty) → PASSES on golden
    try:
        g_count = count_flag_formulas(ws)
        g_total = 45  # rows 2-46
        if g_count == g_total:
            print(f"PASS: Component 2 — All {g_count}/{g_total} Long Wait Flag formulas present (0.20 pts)")
            total_score += 0.20
        elif g_count >= 1:
            c2_partial = round(0.20 * g_count / g_total, 4)
            print(f"PARTIAL: Component 2 — {g_count}/{g_total} Long Wait Flag formulas present ({c2_partial} pts)")
            total_score += c2_partial
        else:
            print(f"FAIL: Component 2 — Expected IF(AND(E?=\"Under Review\",F?>90),\"Follow Up\",\"\") in G2:G46, found {g_count}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Status summary table (I3:J8) with COUNTIF (0.20 points)
    # Requires: I3="Summary by Status", I4:J4 headers, J5:J8 COUNTIF formulas referencing $E$
    # FAILS on initial (I3:J8 are empty) → PASSES on golden
    try:
        i3_val = ws.cell(row=3, column=9).value
        i4_val = ws.cell(row=4, column=9).value
        j4_val = ws.cell(row=4, column=10).value
        status_cf_count = count_status_countif(ws)
        header_present = (i3_val and 'Status' in str(i3_val)
                          and i4_val is not None and j4_val is not None)
        if header_present and status_cf_count == 4:
            print(f"PASS: Component 3 — Status summary in I3:J8 with {status_cf_count}/4 COUNTIF formulas (0.20 pts)")
            total_score += 0.20
        elif status_cf_count >= 1:
            c3_partial = round(0.20 * status_cf_count / 4, 4)
            print(f"PARTIAL: Component 3 — {status_cf_count}/4 status COUNTIF formulas found ({c3_partial} pts)")
            total_score += c3_partial
        else:
            print(f"FAIL: Component 3 — Status summary table missing. I3={repr(i3_val)}, countif={status_cf_count}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Faculty summary table (I11:J57) with COUNTIF (0.20 points)
    # Requires: I11="Summary by Faculty", I12:J12 headers, J13:J57 COUNTIF formulas referencing $A$
    # FAILS on initial (I11:J57 are empty) → PASSES on golden
    try:
        i11_val = ws.cell(row=11, column=9).value
        i12_val = ws.cell(row=12, column=9).value
        j12_val = ws.cell(row=12, column=10).value
        faculty_cf_count = count_faculty_countif(ws)
        header_present = (i11_val and 'Faculty' in str(i11_val)
                          and i12_val is not None and j12_val is not None)
        if header_present and faculty_cf_count >= 10:
            print(f"PASS: Component 4 — Faculty summary in I11:J57 with {faculty_cf_count} COUNTIF formulas (0.20 pts)")
            total_score += 0.20
        elif faculty_cf_count >= 1:
            c4_partial = round(0.20 * min(faculty_cf_count, 45) / 45, 4)
            print(f"PARTIAL: Component 4 — {faculty_cf_count} faculty COUNTIF formulas found ({c4_partial} pts)")
            total_score += c4_partial
        else:
            print(f"FAIL: Component 4 — Faculty summary table missing. I11={repr(i11_val)}, countif={faculty_cf_count}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Conditional formatting — yellow fill on rows where G="Follow Up" (0.10 points)
    # Requires: expression formula referencing G column with "Follow Up" with a fill color
    # FAILS on initial (no CF rules) → PASSES on golden
    try:
        c5_score = score_conditional_formatting(ws)
        if c5_score >= 0.10:
            print(f"PASS: Component 5 — Conditional formatting for 'Follow Up' rows with fill color found ({c5_score} pts)")
            total_score += c5_score
        elif c5_score >= 0.05:
            print(f"PARTIAL: Component 5 — Conditional formatting rule found but fill color not confirmed ({c5_score} pts)")
            total_score += c5_score
        else:
            print(f"FAIL: Component 5 — No conditional formatting for 'Follow Up' rows found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
