"""
Reward Script: Client project portfolio dashboard with status, budget, timeline tracking
Task ID: calc_gpm_083
Domain: libreoffice_calc
Scoring:
  Component 1: Formulas in H4:H11 (% Spent = G/F) — 0.15
  Component 2: Formulas in I4:I11 (Schedule Status) — 0.10
  Component 3: Formulas in J4:J11 (Budget Status) — 0.10
  Component 4: Formulas in K4:K11 (Health) and L4:L11 (RAG) — 0.10
  Component 5: Number formats (H=%, D:E=MMM DD, F:G=$#,##0) — 0.10
  Component 6: Conditional formatting on L column (Red/Amber/Green) — 0.10
  Component 7: Data bars on H4:H11 — 0.05
  Component 8: Conditional formatting on I and J columns — 0.05
  Component 9: Portfolio Summary row 13-14 — 0.10
  Component 10: Two charts (BarChart + PieChart) — 0.15
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_083'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists
    if 'Portfolio' not in wb.sheetnames:
        print("CRITICAL: 'Portfolio' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Portfolio']

    # Component 1: % Spent formulas in H4:H11 (0.15 points)
    # Initial has H4:H11 as None; golden has =G/F formulas
    try:
        formula_count = 0
        for row in range(4, 12):
            val = ws.cell(row=row, column=8).value  # column H
            if val is not None and isinstance(val, str) and '=' in val:
                v_upper = val.upper().replace(' ', '')
                # Should reference G and F in the same row
                if f'G{row}' in v_upper and f'F{row}' in v_upper:
                    formula_count += 1
                elif 'G' in v_upper and 'F' in v_upper:
                    # Looser match: any G/F reference
                    formula_count += 1
        if formula_count >= 7:
            print(f"PASS: Component 1 -- % Spent formulas found in {formula_count}/8 rows (0.15 pts)")
            total_score += 0.15
        elif formula_count >= 4:
            partial = 0.08
            print(f"PARTIAL: Component 1 -- % Spent formulas found in {formula_count}/8 rows ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 -- % Spent formulas found in {formula_count}/8 rows")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Schedule Status formulas in I4:I11 (0.10 points)
    # Initial has I4:I11 as None; golden has IF formulas with TODAY/Overdue/At Risk/On Track
    try:
        formula_count = 0
        for row in range(4, 12):
            val = ws.cell(row=row, column=9).value  # column I
            if val is not None and isinstance(val, str):
                v_upper = val.upper().replace(' ', '')
                if 'IF' in v_upper and ('OVERDUE' in v_upper or 'ONTRACK' in v_upper or 'ATRISK' in v_upper):
                    formula_count += 1
        if formula_count >= 7:
            print(f"PASS: Component 2 -- Schedule Status formulas found in {formula_count}/8 rows (0.10 pts)")
            total_score += 0.10
        elif formula_count >= 4:
            partial = 0.05
            print(f"PARTIAL: Component 2 -- Schedule Status formulas found in {formula_count}/8 rows ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- Schedule Status formulas found in {formula_count}/8 rows")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Budget Status formulas in J4:J11 (0.10 points)
    # Initial has J4:J11 as None; golden has IF formulas with Over Budget/At Risk/On Track
    try:
        formula_count = 0
        for row in range(4, 12):
            val = ws.cell(row=row, column=10).value  # column J
            if val is not None and isinstance(val, str):
                v_upper = val.upper().replace(' ', '')
                if 'IF' in v_upper and ('OVERBUDGET' in v_upper or 'ONTRACK' in v_upper or 'ATRISK' in v_upper):
                    formula_count += 1
        if formula_count >= 7:
            print(f"PASS: Component 3 -- Budget Status formulas found in {formula_count}/8 rows (0.10 pts)")
            total_score += 0.10
        elif formula_count >= 4:
            partial = 0.05
            print(f"PARTIAL: Component 3 -- Budget Status formulas found in {formula_count}/8 rows ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 -- Budget Status formulas found in {formula_count}/8 rows")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Health (K) and RAG (L) formulas in rows 4-11 (0.10 points)
    # Initial has K4:K11 and L4:L11 as None; golden has formulas
    try:
        k_count = 0
        l_count = 0
        for row in range(4, 12):
            k_val = ws.cell(row=row, column=11).value  # column K
            l_val = ws.cell(row=row, column=12).value  # column L
            if k_val is not None and isinstance(k_val, str) and 'IF' in k_val.upper():
                k_count += 1
            if l_val is not None and isinstance(l_val, str) and 'IF' in l_val.upper():
                l_count += 1
        if k_count >= 7 and l_count >= 7:
            print(f"PASS: Component 4 -- Health formulas: K={k_count}/8, L(RAG)={l_count}/8 (0.10 pts)")
            total_score += 0.10
        elif k_count >= 4 or l_count >= 4:
            partial = 0.05
            print(f"PARTIAL: Component 4 -- Health formulas: K={k_count}/8, L(RAG)={l_count}/8 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 -- Health formulas: K={k_count}/8, L(RAG)={l_count}/8")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # Component 5: Number format on H column (% format) (0.10 points)
    # Initial has H4:H11 as None with General format; golden has =G/F with % format
    # NOTE: D:E date format and F:G currency are preconditions (present in initial), NOT scored
    try:
        h_pct_count = 0
        for row in range(4, 12):
            h_fmt = ws.cell(row=row, column=8).number_format
            if h_fmt is not None and '%' in str(h_fmt):
                h_pct_count += 1

        if h_pct_count >= 7:
            print(f"PASS: Component 5 -- H column % format found in {h_pct_count}/8 cells (0.10 pts)")
            total_score += 0.10
        elif h_pct_count >= 4:
            partial = 0.05
            print(f"PARTIAL: Component 5 -- H column % format found in {h_pct_count}/8 cells ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 -- H column % format found in {h_pct_count}/8 cells")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    # Component 6: Conditional formatting on L column - Red/Amber/Green (0.10 points)
    # Initial has no conditional formatting; golden has 4 CF rule sets
    try:
        cf_l_found = False
        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            if 'L' in cf_range:
                for rule in cf.rules:
                    if rule.type == 'cellIs':
                        cf_l_found = True
                        break
            if cf_l_found:
                break
        if cf_l_found:
            print(f"PASS: Component 6 -- Conditional formatting on L column found (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 6 -- No conditional formatting on L column")
    except Exception as e:
        print(f"ERROR: Component 6 -- {e}")

    # Component 7: Data bars on H4:H11 (0.05 points)
    try:
        data_bar_found = False
        for cf in ws.conditional_formatting:
            for rule in cf.rules:
                if rule.type == 'dataBar':
                    cf_range = str(cf)
                    if 'H' in cf_range:
                        data_bar_found = True
                        break
            if data_bar_found:
                break
        if data_bar_found:
            print(f"PASS: Component 7 -- Data bars on H column found (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 7 -- No data bars on H column")
    except Exception as e:
        print(f"ERROR: Component 7 -- {e}")

    # Component 8: Conditional formatting on I and J columns (0.05 points)
    try:
        cf_i_found = False
        cf_j_found = False
        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            for rule in cf.rules:
                if rule.type == 'cellIs':
                    if 'I' in cf_range and not cf_i_found:
                        cf_i_found = True
                    if 'J' in cf_range and not cf_j_found:
                        cf_j_found = True
        if cf_i_found and cf_j_found:
            print(f"PASS: Component 8 -- Conditional formatting on I and J columns found (0.05 pts)")
            total_score += 0.05
        elif cf_i_found or cf_j_found:
            partial = 0.025
            found = 'I' if cf_i_found else 'J'
            print(f"PARTIAL: Component 8 -- CF found on {found} only ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 8 -- No conditional formatting on I or J columns")
    except Exception as e:
        print(f"ERROR: Component 8 -- {e}")

    # Component 9: Portfolio Summary in rows 13-14 (0.10 points)
    # Initial has no data in rows 13+; golden has summary with formulas
    try:
        summary_score = 0.0
        # Check A13 = "Portfolio Summary"
        a13 = ws.cell(row=13, column=1).value
        if a13 is not None and 'portfolio' in str(a13).lower() and 'summary' in str(a13).lower():
            summary_score += 0.03
            print(f"  A13: '{a13}' (Portfolio Summary found)")
        else:
            # Check row 13 and nearby rows for summary label
            found_summary = False
            for r in range(12, 16):
                for c in range(1, 13):
                    v = ws.cell(row=r, column=c).value
                    if v is not None and 'portfolio' in str(v).lower() and 'summary' in str(v).lower():
                        summary_score += 0.03
                        found_summary = True
                        print(f"  {ws.cell(row=r, column=c).coordinate}: '{v}' (Portfolio Summary found)")
                        break
                if found_summary:
                    break
            if not found_summary:
                print(f"  A13: '{a13}' (Portfolio Summary not found)")

        # Check for SUM formulas for total budget/spent
        sum_found = False
        for r in range(12, 16):
            for c in range(1, 16):
                v = ws.cell(row=r, column=c).value
                if v is not None and isinstance(v, str) and 'SUM' in v.upper() and 'F' in v.upper():
                    sum_found = True
                    break
            if sum_found:
                break
        if sum_found:
            summary_score += 0.04
            print(f"  Total budget SUM formula found")
        else:
            print(f"  Total budget SUM formula not found")

        # Check for COUNTIF formulas for RAG counts
        countif_found = False
        for r in range(12, 16):
            for c in range(1, 16):
                v = ws.cell(row=r, column=c).value
                if v is not None and isinstance(v, str) and 'COUNTIF' in v.upper():
                    countif_found = True
                    break
            if countif_found:
                break
        if countif_found:
            summary_score += 0.03
            print(f"  RAG COUNTIF formula found")
        else:
            print(f"  RAG COUNTIF formula not found")

        if summary_score > 0:
            print(f"PASS: Component 9 -- Portfolio Summary ({summary_score:.2f} pts)")
            total_score += summary_score
        else:
            print(f"FAIL: Component 9 -- Portfolio Summary not found")
    except Exception as e:
        print(f"ERROR: Component 9 -- {e}")

    # Component 10: Charts - BarChart + PieChart (0.15 points)
    # Initial has 0 charts; golden has 2
    try:
        charts = ws._charts
        chart_score = 0.0
        bar_found = False
        pie_found = False

        for chart in charts:
            chart_class = chart.__class__.__name__
            # Extract title text
            title_text = ''
            try:
                if chart.title and hasattr(chart.title, 'tx') and chart.title.tx and chart.title.tx.rich:
                    for p in chart.title.tx.rich.p:
                        for r in p.r:
                            title_text += r.t
            except:
                pass

            if chart_class in ('BarChart', 'BarChart3D') or (hasattr(chart, 'type') and chart.type in ('col', 'bar')):
                bar_found = True
                print(f"  Found BarChart: title='{title_text}', series={len(chart.series)}")
            elif chart_class == 'PieChart' or chart_class == 'PieChart3D':
                pie_found = True
                print(f"  Found PieChart: title='{title_text}', series={len(chart.series)}")
            else:
                print(f"  Found chart: class={chart_class}, title='{title_text}'")

        if bar_found and pie_found:
            chart_score = 0.15
            print(f"PASS: Component 10 -- Both BarChart and PieChart found ({chart_score} pts)")
        elif bar_found or pie_found:
            chart_score = 0.08
            which = 'BarChart' if bar_found else 'PieChart'
            print(f"PARTIAL: Component 10 -- Only {which} found ({chart_score} pts)")
        else:
            print(f"FAIL: Component 10 -- No charts found (total charts: {len(charts)})")

        total_score += chart_score
    except Exception as e:
        print(f"ERROR: Component 10 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
