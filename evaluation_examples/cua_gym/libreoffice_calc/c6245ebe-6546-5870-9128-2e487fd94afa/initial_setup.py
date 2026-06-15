"""
Initial Setup: Project Portfolio Dashboard - Q2 2026
Task ID: calc_gpm_083
Domain: libreoffice_calc

Creates the initial spreadsheet with raw project data in Portfolio sheet.
No formulas, no conditional formatting, no charts, no summary row.
"""

import os
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_083'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Portfolio"

    # --- Row 1: Title (merged A1:L1) ---
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = "Project Portfolio Dashboard - Q2 2026"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="FF363636", end_color="FF363636", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Row 3: Headers ---
    headers = [
        "Project", "Client", "PM", "Start", "End",
        "Budget", "Spent", "% Spent", "Schedule Status",
        "Budget Status", "Health", "RAG"
    ]
    charcoal_fill = PatternFill(start_color="FF363636", end_color="FF363636", fill_type="solid")
    white_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = white_font
        cell.fill = charcoal_fill
        cell.alignment = center_align
        cell.border = thin_border

    # --- Rows 4-11: 8 Projects with realistic data ---
    # [Project, Client, PM, Start, End, Budget, Spent]
    projects = [
        ["Website Redesign", "Acme Corp", "Sarah Chen", "2026-01-12", "2026-05-30", 180000, 142500],
        ["Mobile App v3", "TechVision Inc", "Marcus Johnson", "2026-02-01", "2026-07-15", 350000, 210000],
        ["Data Migration", "GlobalBank", "Priya Sharma", "2026-01-05", "2026-03-28", 95000, 98200],
        ["CRM Integration", "Nexus Health", "David Park", "2026-03-01", "2026-06-30", 125000, 78000],
        ["Cloud Infrastructure", "Meridian Logistics", "Elena Rodriguez", "2026-02-15", "2026-08-31", 475000, 195000],
        ["Brand Campaign Portal", "Stellar Media", "James Okonkwo", "2026-01-20", "2026-04-15", 68000, 71400],
        ["Analytics Dashboard", "Vertex Finance", "Anika Patel", "2026-03-10", "2026-09-01", 210000, 52500],
        ["Security Audit Platform", "CyberShield Ltd", "Thomas Wei", "2026-02-20", "2026-05-15", 155000, 130200],
    ]

    from datetime import datetime

    for r_idx, proj in enumerate(projects, 4):
        ws.cell(row=r_idx, column=1, value=proj[0])  # Project
        ws.cell(row=r_idx, column=2, value=proj[1])  # Client
        ws.cell(row=r_idx, column=3, value=proj[2])  # PM
        # Dates as actual date objects
        ws.cell(row=r_idx, column=4, value=datetime.strptime(proj[3], "%Y-%m-%d"))
        ws.cell(row=r_idx, column=5, value=datetime.strptime(proj[4], "%Y-%m-%d"))
        ws.cell(row=r_idx, column=6, value=proj[5])  # Budget
        ws.cell(row=r_idx, column=7, value=proj[6])  # Spent

    # --- Format date columns D:E as MMM DD ---
    for r_idx in range(4, 12):
        ws.cell(row=r_idx, column=4).number_format = 'MMM DD'
        ws.cell(row=r_idx, column=5).number_format = 'MMM DD'

    # --- Format currency columns F:G as $#,##0 ---
    for r_idx in range(4, 12):
        ws.cell(row=r_idx, column=6).number_format = '$#,##0'
        ws.cell(row=r_idx, column=7).number_format = '$#,##0'

    # --- Column widths for readability ---
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 16
    ws.column_dimensions['K'].width = 14
    ws.column_dimensions['L'].width = 10

    # Row 1 height
    ws.row_dimensions[1].height = 36

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
