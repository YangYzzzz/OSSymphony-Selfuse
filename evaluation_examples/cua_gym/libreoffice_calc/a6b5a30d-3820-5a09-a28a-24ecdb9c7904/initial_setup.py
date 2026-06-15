"""
Initial Setup: Build a monthly sales comparison chart with trend lines and formatted data table.
Task ID: calc_gpm_019
Domain: libreoffice_calc

Creates the pre-task spreadsheet with raw, unformatted data only.
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_019'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'MonthlySales'

    # --- Row 1: Headers (plain, no formatting) ---
    ws['A1'] = 'Month'
    ws['B1'] = '2024'
    ws['C1'] = '2025'
    ws['D1'] = 'YoY Growth'

    # --- Rows 2-7: Monthly sales data (raw numbers, no formulas, no formatting) ---
    data = [
        ['Jan', 180000, 205000],
        ['Feb', 165000, 190000],
        ['Mar', 210000, 235000],
        ['Apr', 195000, 228000],
        ['May', 225000, 250000],
        ['Jun', 240000, 262000],
    ]
    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        ws.cell(row=r, column=2, value=row_data[1])
        ws.cell(row=r, column=3, value=row_data[2])
        # D column left empty - task requires adding YoY Growth formulas

    # Set reasonable column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
