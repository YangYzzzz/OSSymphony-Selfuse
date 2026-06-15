"""
Reward Script: Build a monthly sales comparison chart with trend lines and formatted data table.
Task ID: calc_gpm_019
Domain: libreoffice_calc
Scoring:
  Component 1: YoY Growth formulas in D2:D7 with percentage format (0.15)
  Component 2: H1 Total row (row 9) with SUM formulas (0.15)
  Component 3: Header row formatting — bold, centered, dark blue fill, white text (0.15)
  Component 4: Row 9 formatting — bold, double top border, gray fill (0.10)
  Component 5: Currency formatting on B and C columns (0.10)
  Component 6: Line chart present with 2 series and correct title (0.15)
  Component 7: Linear trendlines on both series (0.10)
  Component 8: Conditional formatting with icon set on D2:D7 (0.10)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_019'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists
    if 'MonthlySales' not in wb.sheetnames:
        print("FAIL: Sheet 'MonthlySales' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['MonthlySales']

    # Component 1: YoY Growth formulas in D2:D7 with percentage format (0.15 pts)
    try:
        formula_count = 0
        pct_format_count = 0
        for row in range(2, 8):
            val = ws.cell(row=row, column=4).value
            nf = ws.cell(row=row, column=4).number_format
            if val is not None and isinstance(val, str) and '/' in val and 'C' in val.upper() and 'B' in val.upper():
                formula_count += 1
            if nf and '%' in nf:
                pct_format_count += 1

        if formula_count >= 5 and pct_format_count >= 5:
            print(f"PASS: Component 1 — D2:D7 YoY growth formulas ({formula_count}/6) with pct format ({pct_format_count}/6) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — formulas={formula_count}/6, pct_format={pct_format_count}/6")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: H1 Total row (row 9) with SUM formulas (0.15 pts)
    try:
        a9 = ws.cell(row=9, column=1).value
        b9 = ws.cell(row=9, column=2).value
        c9 = ws.cell(row=9, column=3).value
        d9 = ws.cell(row=9, column=4).value

        has_label = a9 is not None and 'total' in str(a9).lower()
        has_b9_sum = b9 is not None and isinstance(b9, str) and 'SUM' in b9.upper() and 'B2' in b9.upper()
        has_c9_sum = c9 is not None and isinstance(c9, str) and 'SUM' in c9.upper() and 'C2' in c9.upper()
        has_d9_formula = d9 is not None and isinstance(d9, str) and '/' in d9 and ('C9' in d9.upper() or 'B9' in d9.upper())

        checks_passed = sum([has_label, has_b9_sum, has_c9_sum, has_d9_formula])
        if checks_passed >= 3:
            print(f"PASS: Component 2 — Row 9 totals present (label={has_label}, B9_sum={has_b9_sum}, C9_sum={has_c9_sum}, D9_formula={has_d9_formula}) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 — Row 9: label={has_label}, B9={b9!r}, C9={c9!r}, D9={d9!r}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Header row formatting — bold, centered, dark blue fill, white text (0.15 pts)
    try:
        header_pass = 0
        for col in range(1, 5):
            cell = ws.cell(row=1, column=col)
            is_bold = cell.font.bold is True
            is_centered = cell.alignment.horizontal == 'center'
            try:
                fill_rgb = cell.fill.fgColor.rgb if cell.fill.fgColor else None
            except:
                fill_rgb = None
            # Dark blue fill — accept several common dark blue shades
            has_dark_fill = fill_rgb is not None and fill_rgb not in ('00000000', '0',) and cell.fill.patternType == 'solid'
            try:
                font_color = cell.font.color.rgb if cell.font.color else None
            except:
                font_color = None
            has_white_text = font_color is not None and 'FFFFFF' in str(font_color).upper()

            if is_bold and is_centered and has_dark_fill and has_white_text:
                header_pass += 1

        if header_pass >= 3:
            print(f"PASS: Component 3 — Header row formatting correct ({header_pass}/4 cols) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 — Only {header_pass}/4 header cells properly formatted")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Row 9 formatting — bold, double top border, gray fill (0.10 pts)
    try:
        row9_pass = 0
        for col in range(1, 5):
            cell = ws.cell(row=9, column=col)
            is_bold = cell.font.bold is True
            border_top = cell.border.top.style if cell.border.top else None
            has_double_top = border_top == 'double'
            try:
                fill_rgb = cell.fill.fgColor.rgb if cell.fill.fgColor else None
            except:
                fill_rgb = None
            has_gray_fill = fill_rgb is not None and cell.fill.patternType == 'solid' and fill_rgb not in ('00000000', '0',)

            if is_bold and has_double_top and has_gray_fill:
                row9_pass += 1

        if row9_pass >= 3:
            print(f"PASS: Component 4 — Row 9 formatting correct ({row9_pass}/4 cols) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 — Only {row9_pass}/4 row 9 cells properly formatted")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Currency formatting on B and C columns ($#,##0) (0.10 pts)
    try:
        currency_count = 0
        for col in [2, 3]:
            for row in range(2, 8):
                nf = ws.cell(row=row, column=col).number_format
                if nf and '$' in nf:
                    currency_count += 1
        # Also check row 9
        for col in [2, 3]:
            nf = ws.cell(row=9, column=col).number_format
            if nf and '$' in nf:
                currency_count += 1

        if currency_count >= 10:
            print(f"PASS: Component 5 — Currency formatting applied ({currency_count}/14 cells) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 — Currency format found in {currency_count}/14 cells")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Line chart with 2 series and correct title (0.15 pts)
    try:
        charts = ws._charts
        if len(charts) >= 1:
            chart = charts[0]
            is_line = chart.__class__.__name__ == 'LineChart'
            has_2_series = len(chart.series) >= 2

            # Check title
            has_title = False
            if chart.title is not None:
                try:
                    # Extract title text from rich text
                    title_obj = chart.title
                    if hasattr(title_obj, 'tx') and title_obj.tx:
                        if hasattr(title_obj.tx, 'rich') and title_obj.tx.rich:
                            for p in title_obj.tx.rich.p:
                                for r in (p.r or []):
                                    if r.t and '2024' in r.t and '2025' in r.t:
                                        has_title = bool(r.t)
                except:
                    pass

            if is_line and has_2_series and has_title:
                print(f"PASS: Component 6 — Line chart with 2 series and correct title (0.15 pts)")
                total_score += 0.15
            elif is_line and has_2_series:
                print(f"PARTIAL: Component 6 — Line chart with 2 series but title issue (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 6 — is_line={is_line}, series={len(chart.series)}, has_title={has_title}")
        else:
            print(f"FAIL: Component 6 — No charts found")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: Linear trendlines on both chart series (0.10 pts)
    try:
        charts = ws._charts
        if len(charts) >= 1:
            chart = charts[0]
            trendline_count = 0
            for s in chart.series:
                if hasattr(s, 'trendline') and s.trendline is not None:
                    if s.trendline.trendlineType == 'linear':
                        trendline_count += 1

            if trendline_count >= 2:
                print(f"PASS: Component 7 — Both series have linear trendlines (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 7 — Only {trendline_count}/2 series have linear trendlines")
        else:
            print(f"FAIL: Component 7 — No charts found")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    # Component 8: Conditional formatting with icon set on D2:D7 (0.10 pts)
    try:
        has_icon_set = False
        has_cf_on_d = False
        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            if 'D' in cf_range:
                has_cf_on_d = bool(cf_range)
                for rule in cf.rules:
                    if rule.type == 'iconSet':
                        has_icon_set = bool(rule.type)

        if has_icon_set and has_cf_on_d:
            print(f"PASS: Component 8 — Conditional formatting with icon set on D column (0.10 pts)")
            total_score += 0.10
        elif has_cf_on_d:
            print(f"PARTIAL: Component 8 — Conditional formatting on D but no icon set (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 8 — No conditional formatting on D column")
    except Exception as e:
        print(f"ERROR: Component 8 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
