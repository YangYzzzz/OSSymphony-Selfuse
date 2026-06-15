"""
Reward Script: Verify SUMPRODUCT weighted score formulas in Summary!C2:C6
Task ID: calc_mcp_052
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): All 5 cells C2:C6 contain formulas (non-empty strings starting with '=')
  Component 2 (0.3): Each formula uses SUMPRODUCT function
  Component 3 (0.3): Each formula references Scores sheet with category filtering and
                      multiplies Score (col B) by Weight (col C)
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_052'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Summary sheet must exist
    if 'Summary' not in wb.sheetnames:
        print("FAIL: 'Summary' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Summary']

    # Read C2:C6 values
    formulas = {}
    for row in range(2, 7):
        val = ws.cell(row=row, column=3).value
        formulas[row] = val

    print(f"DEBUG: C2:C6 values = {formulas}")

    # Component 1: All 5 cells C2:C6 contain formulas (0.4 points)
    # A formula is a non-empty string starting with '='
    # This FAILS on initial (all None) and PASSES on golden (all have formulas)
    try:
        formula_count = 0
        for row in range(2, 7):
            val = formulas[row]
            if isinstance(val, str) and val.startswith('='):
                formula_count += 1
            else:
                print(f"FAIL: C{row} is not a formula, found: {repr(val)}")

        if formula_count == 5:
            print(f"PASS: Component 1 — All 5 cells C2:C6 contain formulas (0.4 pts)")
            total_score += 0.4
        elif formula_count > 0:
            partial = round(0.4 * (formula_count / 5), 2)
            print(f"PARTIAL: Component 1 — {formula_count}/5 cells have formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No formulas found in C2:C6")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Each formula uses SUMPRODUCT (0.3 points)
    # SUMPRODUCT is the expected function for weighted sums with category filtering
    # This FAILS on initial (no formulas) and PASSES on golden (all SUMPRODUCT)
    try:
        sumproduct_count = 0
        for row in range(2, 7):
            val = formulas[row]
            if isinstance(val, str) and 'SUMPRODUCT' in val.upper():
                sumproduct_count += 1
            else:
                print(f"FAIL: C{row} does not use SUMPRODUCT, found: {repr(val)}")

        if sumproduct_count == 5:
            print(f"PASS: Component 2 — All 5 formulas use SUMPRODUCT (0.3 pts)")
            total_score += 0.3
        elif sumproduct_count > 0:
            partial = round(0.3 * (sumproduct_count / 5), 2)
            print(f"PARTIAL: Component 2 — {sumproduct_count}/5 formulas use SUMPRODUCT ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No SUMPRODUCT formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Each formula correctly references Scores sheet with
    # category filtering (col A comparison) AND multiplies B * C columns (0.3 points)
    # This FAILS on initial (no formulas) and PASSES on golden
    try:
        correct_count = 0
        for row in range(2, 7):
            val = formulas[row]
            if not isinstance(val, str):
                print(f"FAIL: C{row} is not a formula string")
                continue

            upper_val = val.upper().replace(' ', '')

            # Check references to Scores sheet columns A, B, and C
            has_scores_a = bool(re.search(r'SCORES!A', val, re.IGNORECASE))
            has_scores_b = bool(re.search(r'SCORES!B', val, re.IGNORECASE))
            has_scores_c = bool(re.search(r'SCORES!C', val, re.IGNORECASE))
            # Category filtering: formula should compare Scores!A range with a category ref
            has_category_filter = bool(re.search(r'SCORES!A[^)]*=\s*A\d', val, re.IGNORECASE))

            if has_scores_a and has_scores_b and has_scores_c and has_category_filter:
                correct_count += 1
            else:
                print(f"FAIL: C{row} formula missing required references: "
                      f"Scores!A={has_scores_a}, Scores!B={has_scores_b}, "
                      f"Scores!C={has_scores_c}, category_filter={has_category_filter}")
                print(f"  Formula: {val}")

        if correct_count == 5:
            print(f"PASS: Component 3 — All 5 formulas correctly reference Scores sheet with filtering (0.3 pts)")
            total_score += 0.3
        elif correct_count > 0:
            partial = round(0.3 * (correct_count / 5), 2)
            print(f"PARTIAL: Component 3 — {correct_count}/5 formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No formulas have correct Scores sheet references")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
