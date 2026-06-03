"""
Initial Setup: Create spreadsheet with Scores data and Summary sheet for weighted score calculation
Task ID: calc_mcp_052
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_052'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

CATEGORIES = ['Electronics', 'Furniture', 'Clothing', 'Groceries', 'Sports']

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()

    # --- Sheet 1: Scores ---
    ws_scores = wb.active
    ws_scores.title = 'Scores'
    ws_scores['A1'] = 'Category'
    ws_scores['B1'] = 'Score'
    ws_scores['C1'] = 'Weight'

    # Make headers bold
    from openpyxl.styles import Font
    bold = Font(bold=True)
    for col in ['A', 'B', 'C']:
        ws_scores[f'{col}1'].font = bold

    # Fill rows 2-100 with realistic data
    for row_idx in range(2, 101):
        cat = random.choice(CATEGORIES)
        score = round(random.uniform(10, 100), 2)
        weight = round(random.uniform(0.1, 1.0), 2)
        ws_scores.cell(row=row_idx, column=1, value=cat)
        ws_scores.cell(row=row_idx, column=2, value=score)
        ws_scores.cell(row=row_idx, column=3, value=weight)

    # Set column widths
    ws_scores.column_dimensions['A'].width = 15
    ws_scores.column_dimensions['B'].width = 12
    ws_scores.column_dimensions['C'].width = 12

    # --- Sheet 2: Summary ---
    ws_summary = wb.create_sheet('Summary')
    ws_summary['A1'] = 'Category'
    ws_summary['B1'] = 'Count'
    ws_summary['C1'] = 'Weighted Total'

    for col in ['A', 'B', 'C']:
        ws_summary[f'{col}1'].font = bold

    # Fill category names in A2:A6
    for i, cat in enumerate(CATEGORIES, 2):
        ws_summary.cell(row=i, column=1, value=cat)

    # Fill Count column (B2:B6) with COUNTIF formulas for context
    for i in range(2, 7):
        ws_summary.cell(row=i, column=2, value=f"=COUNTIF(Scores!A$2:A$100,A{i})")

    # C2:C6 intentionally left EMPTY - this is the task target

    ws_summary.column_dimensions['A'].width = 15
    ws_summary.column_dimensions['B'].width = 12
    ws_summary.column_dimensions['C'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
