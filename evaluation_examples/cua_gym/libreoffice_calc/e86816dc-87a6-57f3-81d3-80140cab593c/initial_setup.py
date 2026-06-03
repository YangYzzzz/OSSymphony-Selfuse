"""
Initial Setup: Fill employee badge number column with sequential IDs
Task ID: osworld_calc_fill_sequence_numbers_007
Domain: libreoffice_calc

Creates a spreadsheet with employee records where column A (Badge Number) is empty.
The agent must fill column A with IDs in format 'EMP-[Dept3LetterCode]-[3-digit number]'.
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_fill_sequence_numbers_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Employees ---
    ws = wb.active
    ws.title = 'Employees'

    # Headers: Badge Number (A, empty for task), Department (B), Name (C), Role (D)
    headers = ['Badge Number', 'Department', 'Name', 'Role']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Employee data — realistic content across multiple departments
    # Badge Number column (A) is intentionally LEFT EMPTY (task is to fill it)
    # Ordered by department to make sequential numbering clear
    # Each tuple: (Department, Name, Role) — Badge Number column stays None/empty
    data = [
        # Department,         Name,                   Role
        ('Finance',           'Sarah Chen',           'Financial Analyst'),
        ('Finance',           'Marcus Johnson',       'Senior Accountant'),
        ('Finance',           'Linda Perez',          'Budget Manager'),
        ('Operations',        'James Wright',         'Operations Lead'),
        ('Operations',        'Aisha Patel',          'Process Engineer'),
        ('Operations',        'Daniel Kim',           'Supply Chain Analyst'),
        ('Operations',        'Rachel Torres',        'Logistics Coordinator'),
        ('Marketing',         'Emily Nguyen',         'Marketing Manager'),
        ('Marketing',         'Tyler Brooks',         'Content Strategist'),
        ('Marketing',         'Sofia Reyes',          'Brand Specialist'),
        ('Engineering',       'Nathan Okafor',        'Software Engineer'),
        ('Engineering',       'Priya Sharma',         'DevOps Engineer'),
        ('Engineering',       'Lucas Hoffmann',       'Backend Developer'),
        ('Engineering',       'Chloe Anderson',       'QA Engineer'),
        ('Accounting',        'Omar Hassan',          'Staff Accountant'),
        ('Accounting',        'Vanessa Liu',          'Payroll Specialist'),
        ('Human Resources',   'Ryan Murphy',          'HR Business Partner'),
        ('Human Resources',   'Diana Kowalski',       'Recruitment Specialist'),
        ('Human Resources',   'Andre Williams',       'Training Coordinator'),
    ]

    for r, (dept, name, role) in enumerate(data, 2):
        # Column A: leave empty (None) — Badge Number is what the agent must fill
        ws.cell(row=r, column=1, value=None)
        ws.cell(row=r, column=2, value=dept)
        ws.cell(row=r, column=3, value=name)
        ws.cell(row=r, column=4, value=role)

    # Column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 28

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
