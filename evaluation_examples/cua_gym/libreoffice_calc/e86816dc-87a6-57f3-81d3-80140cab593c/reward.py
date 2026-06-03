"""
Reward Script: Fill employee badge number column with sequential IDs
Task ID: osworld_calc_fill_sequence_numbers_007
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.4): All badge number cells are filled (no None/empty in column A, rows 2-20)
  - Component 2 (0.3): All values match the EMP-[3-LETTER-CODE]-[3-digit-seq] format
  - Component 3 (0.3): Department codes are first-3-char uppercase abbreviations and sequential
                       numbering within each department is correct (001, 002, ...)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_fill_sequence_numbers_007'

def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Employees' sheet must exist
    if 'Employees' not in wb.sheetnames:
        print("CRITICAL: 'Employees' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Employees']

    # Collect badge numbers and department values from rows 2-20 (data rows)
    badge_numbers = []
    departments = []
    try:
        for row in range(2, 21):  # rows 2 through 20 (19 employees)
            badge_val = ws.cell(row=row, column=1).value
            dept_val = ws.cell(row=row, column=2).value
            badge_numbers.append(badge_val)
            departments.append(dept_val)
    except Exception as e:
        print(f"ERROR: Could not read employee rows: {e}")
        print("REWARD: 0.0")
        return 0.0

    # --- Component 1: All badge number cells are non-empty (0.4 points) ---
    # This checks that column A rows 2-20 are no longer None/empty
    try:
        filled_count = sum(1 for b in badge_numbers if b is not None and str(b).strip() != '')
        total_cells = len(badge_numbers)  # should be 19
        if filled_count == total_cells:
            print(f"PASS: Component 1 — All {total_cells} badge number cells are filled (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — Only {filled_count}/{total_cells} badge number cells are filled")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # --- Component 2: All values match EMP-[3-LETTER]-[3-digit] format (0.3 points) ---
    # Pattern: EMP-XXX-NNN where X is uppercase letter, N is digit
    try:
        badge_pattern = re.compile(r'^EMP-[A-Z]{3}-\d{3}$')
        invalid_badges = []
        for i, badge in enumerate(badge_numbers):
            if badge is None or not badge_pattern.match(str(badge)):
                invalid_badges.append((i + 2, badge))  # row number + value
        if len(invalid_badges) == 0:
            print(f"PASS: Component 2 — All badge numbers match EMP-[3-LETTER]-[3-digit] format (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — {len(invalid_badges)} badge numbers have invalid format:")
            for row_num, val in invalid_badges[:5]:  # show up to 5 errors
                print(f"  Row {row_num}: '{val}'")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # --- Component 3: Department codes are consistent per dept, and sequential numbering is correct (0.3 points) ---
    # Rules:
    #   - All employees in the same department must share the same 3-letter dept code
    #   - Sequential numbering within each department must start at 1 and increment by 1
    #     (i.e., 001, 002, 003, ... in order of appearance)
    # Note: The exact abbreviation used (e.g., OPS vs OPE for Operations) is up to the agent;
    #       we only verify consistency and correct sequential ordering within each department group.
    try:
        errors = []

        # Build a mapping: dept_name -> list of (seq_int, row_num) in appearance order
        dept_seq_map = {}  # dept_name -> ordered list of actual sequence numbers
        dept_code_map = {}  # dept_name -> set of dept codes used

        for i, (badge, dept) in enumerate(zip(badge_numbers, departments)):
            row_num = i + 2
            if badge is None or dept is None:
                errors.append(f"Row {row_num}: badge or dept is None")
                continue

            badge_str = str(badge).strip()
            dept_str = str(dept).strip()

            parts = badge_str.split('-')
            if len(parts) != 3 or parts[0] != 'EMP':
                errors.append(f"Row {row_num}: badge '{badge_str}' has wrong structure")
                continue

            actual_dept_code = parts[1]
            actual_seq_str = parts[2]

            # Track the dept code used for this department name (must be consistent)
            if dept_str not in dept_code_map:
                dept_code_map[dept_str] = set()
            dept_code_map[dept_str].add(actual_dept_code)

            # Track the sequence number
            try:
                actual_seq = int(actual_seq_str)
            except ValueError:
                errors.append(f"Row {row_num}: non-numeric seq '{actual_seq_str}'")
                continue

            if dept_str not in dept_seq_map:
                dept_seq_map[dept_str] = []
            dept_seq_map[dept_str].append((actual_seq, row_num))

        # Check consistency: each department name must map to exactly one dept code
        for dept_name, codes in dept_code_map.items():
            if len(codes) > 1:
                errors.append(f"Dept '{dept_name}' uses multiple codes: {codes}")

        # Check sequential numbering: within each department, sequence must be 1,2,3,...
        for dept_name, seq_list in dept_seq_map.items():
            for idx, (actual_seq, row_num) in enumerate(seq_list):
                expected_seq = idx + 1
                if actual_seq != expected_seq:
                    errors.append(
                        f"Row {row_num} (dept '{dept_name}'): seq {actual_seq} != expected {expected_seq}"
                    )

        if len(errors) == 0:
            print(f"PASS: Component 3 — Dept codes consistent per department, sequential numbering correct (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 3 — {len(errors)} errors in dept code consistency / sequential numbering:")
            for err in errors[:5]:  # show up to 5 errors
                print(f"  {err}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in a given env
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
