"""
Reward Script: Nested VLOOKUP formula in LibreOffice Calc
Task ID: calc_lf_018
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): C2 contains a formula (not empty/literal)
  Component 2 (0.4): C2 contains a nested VLOOKUP (VLOOKUP inside VLOOKUP)
  Component 3 (0.3): Formula references the Budgets sheet for the outer lookup
"""

import os
import re
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_018'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Staff sheet must exist
    if 'Staff' not in wb.sheetnames:
        print("FAIL: 'Staff' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Staff']

    # Get C2 value (formula string or literal)
    c2_val = ws['C2'].value

    # Component 1: C2 contains a formula (0.3 points)
    # In the initial file C2 is None. A formula starts with '='.
    try:
        if c2_val is not None and isinstance(c2_val, str) and c2_val.strip().startswith('='):
            print(f"PASS: Component 1 - C2 contains a formula: {c2_val!r} (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 - C2 does not contain a formula, found: {c2_val!r}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Formula uses nested VLOOKUP (VLOOKUP inside VLOOKUP) (0.4 points)
    # We check that the formula contains at least two VLOOKUP calls, one nested inside the other.
    try:
        if c2_val and isinstance(c2_val, str):
            formula_upper = c2_val.upper().replace(" ", "")
            # Count VLOOKUP occurrences
            vlookup_count = formula_upper.count('VLOOKUP(')
            if vlookup_count >= 2:
                print(f"PASS: Component 2 - Nested VLOOKUP detected ({vlookup_count} VLOOKUP calls) (0.4 pts)")
                total_score += 0.4
            else:
                print(f"FAIL: Component 2 - Expected nested VLOOKUP (>=2 calls), found {vlookup_count}")
        else:
            print(f"FAIL: Component 2 - C2 is not a formula string")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Outer VLOOKUP references the Budgets sheet (0.3 points)
    # The formula should reference Budgets sheet data (e.g., Budgets.A2:B4 or Budgets!A2:B4)
    # to look up the department budget from the Budgets table.
    try:
        if c2_val and isinstance(c2_val, str):
            formula_upper = c2_val.upper().replace(" ", "")
            # Check for reference to Budgets sheet (LibreOffice uses '.' separator, Excel uses '!')
            if 'BUDGETS.' in formula_upper or 'BUDGETS!' in formula_upper:
                print(f"PASS: Component 3 - Formula references Budgets sheet (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 3 - Formula does not reference 'Budgets' sheet. Formula: {c2_val!r}")
        else:
            print(f"FAIL: Component 3 - C2 is not a formula string")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
