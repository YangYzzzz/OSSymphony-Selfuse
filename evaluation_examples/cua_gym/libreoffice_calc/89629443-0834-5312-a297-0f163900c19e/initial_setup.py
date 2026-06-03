"""
Initial Setup: Create a spreadsheet with Staff and Budgets sheets for nested VLOOKUP task.
Task ID: calc_lf_018
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_018'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 'Staff' ---
    ws_staff = wb.active
    ws_staff.title = 'Staff'

    # Headers
    ws_staff['A1'] = 'Name'
    ws_staff['B1'] = 'Dept'
    ws_staff['C1'] = 'Dept Budget'

    # Data
    staff_data = [
        ['Alice', 'IT'],
        ['Bob', 'HR'],
        ['Carol', 'Sales'],
    ]
    for r, row_data in enumerate(staff_data, 2):
        ws_staff.cell(row=r, column=1, value=row_data[0])
        ws_staff.cell(row=r, column=2, value=row_data[1])

    # C2 is intentionally left EMPTY - the task asks the agent to enter a formula there

    # --- Sheet 'Budgets' ---
    ws_budgets = wb.create_sheet('Budgets')

    ws_budgets['A1'] = 'Dept'
    ws_budgets['B1'] = 'Budget'

    budget_data = [
        ['HR', 50000],
        ['IT', 75000],
        ['Sales', 90000],
    ]
    for r, row_data in enumerate(budget_data, 2):
        ws_budgets.cell(row=r, column=1, value=row_data[0])
        ws_budgets.cell(row=r, column=2, value=row_data[1])

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Open in LibreOffice Calc for GUI-ready state
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
