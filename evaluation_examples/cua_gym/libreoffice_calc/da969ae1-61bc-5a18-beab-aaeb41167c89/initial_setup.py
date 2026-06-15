"""
Initial Setup: Fill A2:A22 with percentages from 0 to 100 in steps of 5
Task ID: calc_dop_fillseries_linear_075
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_dop_fillseries_linear_075'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'PercentageTable'

    # Row 1 headers
    ws['A1'] = 'Percentage'
    ws['B1'] = 'Decimal'
    ws['C1'] = 'Description'

    # A2: first percentage value (0) — already entered
    ws['A2'] = 0

    # A3:A22: EMPTY — to be filled by the agent using Fill Series

    # B2: decimal formula (already entered)
    ws['B2'] = '=A2/100'

    # B3:B22: EMPTY — to be filled down by the agent

    # Column C: 21 description labels already filled in (rows 2–22)
    descriptions = [
        'Negligible',
        'Very Low',
        'Low',
        'Low-Moderate',
        'Moderate',
        'Moderate',
        'Moderate-High',
        'High',
        'High',
        'High',
        'Very High',
        'Very High',
        'Very High',
        'Near Complete',
        'Near Complete',
        'Near Complete',
        'Near Complete',
        'Almost Full',
        'Almost Full',
        'Full',
        'Complete',
    ]
    for i, desc in enumerate(descriptions, start=2):
        ws.cell(row=i, column=3, value=desc)

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
