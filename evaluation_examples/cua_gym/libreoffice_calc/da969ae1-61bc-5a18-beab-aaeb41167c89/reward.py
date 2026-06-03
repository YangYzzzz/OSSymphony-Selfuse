"""
Reward Script: Fill A2:A22 with percentages from 0 to 100 in steps of 5 and fill down B column formulas
Task ID: calc_dop_fillseries_linear_075
Domain: libreoffice_calc
Scoring:
  Component 1 (0.6 pts): Column A (A3:A22) filled with linear series 5, 10, 15, ..., 100
  Component 2 (0.4 pts): Column B (B3:B22) filled down with =An/100 formulas
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_dop_fillseries_linear_075'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet 'PercentageTable' exists (precondition gate)
    if 'PercentageTable' not in wb.sheetnames:
        print("FAIL: Sheet 'PercentageTable' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['PercentageTable']

    # Component 1: Column A (A3:A22) filled with linear series 5, 10, 15, ..., 100 (0.6 points)
    # In the initial file, A3:A22 are all None; in the golden file they are 5, 10, ..., 100
    # We verify all 20 values are present and correct (step=5, starting from 5 in A3)
    try:
        expected_a_values = list(range(5, 105, 5))  # [5, 10, 15, ..., 100]
        actual_a_values = []
        for row in range(3, 23):  # rows 3 through 22 (A3:A22)
            val = ws.cell(row=row, column=1).value
            actual_a_values.append(val)

        # Check if all values match expected linear series
        all_correct = True
        wrong_cells = []
        for i, (exp, act) in enumerate(zip(expected_a_values, actual_a_values)):
            row_num = i + 3
            if act is None or act != exp:
                all_correct = False
                wrong_cells.append(f"A{row_num}: expected {exp}, got {repr(act)}")

        if all_correct:
            print(f"PASS: Component 1 — A3:A22 filled with linear series 5..100 in steps of 5 (0.6 pts)")
            total_score += 0.6
        else:
            # Partial credit: count how many cells are correct
            correct_count = sum(
                1 for exp, act in zip(expected_a_values, actual_a_values)
                if act is not None and act == exp
            )
            if correct_count > 0:
                partial = round(0.6 * correct_count / 20, 2)
                print(f"PARTIAL: Component 1 — {correct_count}/20 cells correct in A3:A22 ({partial} pts)")
                print(f"  First wrong cells: {wrong_cells[:3]}")
                total_score += partial
            else:
                print(f"FAIL: Component 1 — A3:A22 not filled. First cells: {actual_a_values[:3]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Column B (B3:B22) filled with formulas =An/100 (0.4 points)
    # In the initial file, B3:B22 are all None; in the golden file they are =A3/100, =A4/100, etc.
    try:
        all_formulas_correct = True
        wrong_b_cells = []
        for row in range(3, 23):  # rows 3 through 22 (B3:B22)
            val = ws.cell(row=row, column=2).value
            expected_formula = f"=A{row}/100"
            # Normalize comparison: strip spaces, uppercase
            if val is None:
                all_formulas_correct = False
                wrong_b_cells.append(f"B{row}: expected '{expected_formula}', got None")
            elif not (isinstance(val, str) and val.strip().upper().replace(" ", "") == expected_formula.upper().replace(" ", "")):
                all_formulas_correct = False
                wrong_b_cells.append(f"B{row}: expected '{expected_formula}', got '{val}'")

        if all_formulas_correct:
            print(f"PASS: Component 2 — B3:B22 filled with =An/100 formulas (0.4 pts)")
            total_score += 0.4
        else:
            # Partial credit: count how many formulas are correct
            correct_count = 0
            for row in range(3, 23):
                val = ws.cell(row=row, column=2).value
                expected_formula = f"=A{row}/100"
                if val is not None and isinstance(val, str) and val.strip().upper().replace(" ", "") == expected_formula.upper().replace(" ", ""):
                    correct_count += 1

            if correct_count > 0:
                partial = round(0.4 * correct_count / 20, 2)
                print(f"PARTIAL: Component 2 — {correct_count}/20 B formulas correct ({partial} pts)")
                print(f"  First wrong cells: {wrong_b_cells[:3]}")
                total_score += partial
            else:
                print(f"FAIL: Component 2 — B3:B22 not filled with formulas.")
                print(f"  First wrong cells: {wrong_b_cells[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {round(final_score, 4)}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
