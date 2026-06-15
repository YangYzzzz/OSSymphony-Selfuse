"""
Reward Script: Modify named range 'Revenue' from B2:B13 to B2:B15
Task ID: calc_nrv_012
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): Named range 'Revenue' end row is B15 (the key change)
  Component 2 (0.3): Named range 'Revenue' start row is B2 (correct start preserved)
  Component 3 (0.2): Named range references correct sheet and data B14/B15 intact
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_012'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify that the named range 'Revenue' has been modified from $B$2:$B$13 to $B$2:$B$15.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Locate the 'Revenue' named range
    revenue_range = None
    try:
        for name, dn in wb.defined_names.items():
            if name.lower() == 'revenue':
                revenue_range = dn.attr_text
                break
    except Exception as e:
        print(f"ERROR: Could not iterate defined names: {e}")

    if revenue_range is None:
        print("FAIL: Named range 'Revenue' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    print(f"INFO: Found named range 'Revenue' = {revenue_range}")

    # Parse the range string, e.g. "'Revenue Data'!$B$2:$B$15"
    # Normalize by removing quotes, dollar signs, spaces
    range_upper = revenue_range.upper().replace("'", "").replace(" ", "")

    # Parse start and end rows from the range string
    end_row = None
    start_row = None
    try:
        end_match = re.search(r'\$?B\$?(\d+)\s*$', range_upper)
        if end_match:
            end_row = int(end_match.group(1))
        start_match = re.search(r'!\$?B\$?(\d+):', range_upper)
        if start_match:
            start_row = int(start_match.group(1))
    except Exception as e:
        print(f"ERROR: Could not parse range: {e}")

    # Component 1: Named range end row is B15 (0.5 points)
    # This is THE task-introduced change: extending from B13 to B15
    try:
        if end_row == 15:
            print(f"PASS: Component 1 — Named range ends at B15 (0.5 pts)")
            total_score += 0.5
        elif end_row is not None:
            print(f"FAIL: Component 1 — Named range ends at B{end_row}, expected B15")
        else:
            print(f"FAIL: Component 1 — Could not parse end row from: {revenue_range}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Full range is exactly B2:B15 (0.3 points)
    # Anchored to the change: end must be B15 AND start must be B2
    # This fails on initial_env because end_row=13, not 15
    try:
        if end_row == 15 and start_row == 2:
            print(f"PASS: Component 2 — Named range is B2:B15 (correct full range) (0.3 pts)")
            total_score += 0.3
        elif end_row != 15:
            print(f"FAIL: Component 2 — End row is B{end_row}, not B15 (range not extended)")
        elif start_row != 2:
            print(f"FAIL: Component 2 — Start row is B{start_row}, expected B2")
        else:
            print(f"FAIL: Component 2 — Could not verify range")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Range ends at B15 AND data in B14/B15 is intact (0.2 points)
    # Compound check: range change + data integrity for newly included cells
    # Fails on initial_env because end_row != 15
    try:
        ws = wb.active
        b14_val = ws.cell(row=14, column=2).value
        b15_val = ws.cell(row=15, column=2).value
        data_ok = (b14_val == 7200 and b15_val == 7500)

        if end_row == 15 and data_ok:
            print(f"PASS: Component 3 — Range includes B15 and data intact (B14={b14_val}, B15={b15_val}) (0.2 pts)")
            total_score += 0.2
        elif end_row != 15:
            print(f"FAIL: Component 3 — Range not extended to B15 (end_row=B{end_row})")
        else:
            print(f"FAIL: Component 3 — Data integrity: B14={b14_val} (expected 7200), B15={b15_val} (expected 7500)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
