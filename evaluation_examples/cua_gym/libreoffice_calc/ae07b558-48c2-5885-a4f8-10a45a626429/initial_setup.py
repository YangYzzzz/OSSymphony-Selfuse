"""
Initial Setup: Named range 'Revenue' pointing to B2:B13 with 15 rows of data
Task ID: calc_nrv_012
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.workbook.defined_name import DefinedName

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_012'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Revenue Data ---
    ws = wb.active
    ws.title = 'Revenue Data'

    # Headers
    ws.cell(row=1, column=1, value='Month')
    ws.cell(row=1, column=2, value='Revenue ($)')

    # 12 months of actual data (B2:B13)
    monthly_data = [
        ('Jan 2025', 5800),
        ('Feb 2025', 6100),
        ('Mar 2025', 6450),
        ('Apr 2025', 5950),
        ('May 2025', 6300),
        ('Jun 2025', 6750),
        ('Jul 2025', 7100),
        ('Aug 2025', 6800),
        ('Sep 2025', 6500),
        ('Oct 2025', 6900),
        ('Nov 2025', 7050),
        ('Dec 2025', 7300),
    ]

    for r, (month, revenue) in enumerate(monthly_data, 2):
        ws.cell(row=r, column=1, value=month)
        ws.cell(row=r, column=2, value=revenue)

    # 3 forecast extension months (B14:B15 per task; B14=7200, B15=7500)
    # Task context says "3 extra months" but specifies B14 and B15 (2 extra rows of data)
    # B14 and B15 are the newly added forecast months
    ws.cell(row=14, column=1, value='Jan 2026 (Forecast)')
    ws.cell(row=14, column=2, value=7200)
    ws.cell(row=15, column=1, value='Feb 2026 (Forecast)')
    ws.cell(row=15, column=2, value=7500)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 15

    # Define named range 'Revenue' pointing to $B$2:$B$13 (initial state)
    named_range = DefinedName('Revenue', attr_text="'Revenue Data'!$B$2:$B$13")
    wb.defined_names.add(named_range)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
