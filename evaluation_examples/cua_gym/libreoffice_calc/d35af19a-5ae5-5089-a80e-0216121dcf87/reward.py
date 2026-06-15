"""
Reward Script: Build a sales headcount and capacity planning model
Task ID: calc_sales_headcount_planning_070
Domain: libreoffice_calc
Scoring:
  - Component 1: Gap-to-fill formula in B4 (0.25 pts)
  - Component 2: Rep calculation formulas in B7 and B8 (0.25 pts)
  - Component 3: B7 and B8 highlighted with bold + yellow background (0.25 pts)
  - Component 4: Quarterly hiring plan filled in B12:E16 with H1-weighted distribution (0.25 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_headcount_planning_070'

def normalize_formula(formula):
    """Normalize formula for comparison: upper-case, remove spaces."""
    if not isinstance(formula, str):
        return ''
    return formula.upper().replace(' ', '').strip()


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'HeadcountPlan' not in wb.sheetnames:
        print("CRITICAL: Sheet 'HeadcountPlan' not found.")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['HeadcountPlan']

    # -----------------------------------------------------------------------
    # Component 1: Gap-to-fill formula in B4 (0.25 pts)
    # B4 should contain =B2-B3 formula. In the initial file B4 is empty.
    # -----------------------------------------------------------------------
    try:
        b4_val = ws.cell(row=4, column=2).value
        b4_norm = normalize_formula(b4_val)
        expected_b4 = '=B2-B3'
        if b4_norm == normalize_formula(expected_b4):
            print(f"PASS: Component 1 — B4 contains gap-to-fill formula '{b4_val}' (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — B4 expected formula '{expected_b4}', found: {repr(b4_val)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Rep calculation formulas in B7 and B8 (0.25 pts)
    # B7 should contain =CEILING(B4/B5,1)
    # B8 should contain =CEILING(B4/(B5*B6),1)
    # Both empty in initial file.
    # -----------------------------------------------------------------------
    try:
        b7_val = ws.cell(row=7, column=2).value
        b8_val = ws.cell(row=8, column=2).value
        b7_norm = normalize_formula(b7_val)
        b8_norm = normalize_formula(b8_val)
        expected_b7 = '=CEILING(B4/B5,1)'
        expected_b8 = '=CEILING(B4/(B5*B6),1)'

        b7_ok = b7_norm == normalize_formula(expected_b7)
        b8_ok = b8_norm == normalize_formula(expected_b8)

        if b7_ok and b8_ok:
            print(f"PASS: Component 2 — B7='{b7_val}', B8='{b8_val}' — both rep formulas correct (0.25 pts)")
            total_score += 0.25
        elif b7_ok:
            print(f"PARTIAL: Component 2 — B7 formula OK, but B8 wrong. Expected '{expected_b8}', got: {repr(b8_val)}")
        elif b8_ok:
            print(f"PARTIAL: Component 2 — B8 formula OK, but B7 wrong. Expected '{expected_b7}', got: {repr(b7_val)}")
        else:
            print(f"FAIL: Component 2 — B7 expected '{expected_b7}' got {repr(b7_val)}; B8 expected '{expected_b8}' got {repr(b8_val)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: B7 and B8 bold + yellow highlight (0.25 pts)
    # In the initial file, B7 and B8 are empty and have no special formatting.
    # Golden has bold=True and fgColor=FFFFFF00 (yellow) for both.
    # -----------------------------------------------------------------------
    try:
        b7_cell = ws.cell(row=7, column=2)
        b8_cell = ws.cell(row=8, column=2)

        b7_bold = b7_cell.font.bold is True
        b8_bold = b8_cell.font.bold is True

        try:
            b7_fill = b7_cell.fill.fgColor.rgb
            b8_fill = b8_cell.fill.fgColor.rgb
        except Exception:
            b7_fill = None
            b8_fill = None

        # Yellow: FFFFFF00 (8-char ARGB)
        YELLOW = 'FFFFFF00'
        b7_yellow = (b7_fill is not None and b7_fill.upper() == YELLOW)
        b8_yellow = (b8_fill is not None and b8_fill.upper() == YELLOW)

        b7_formatted = b7_bold and b7_yellow
        b8_formatted = b8_bold and b8_yellow

        if b7_formatted and b8_formatted:
            print(f"PASS: Component 3 — B7 and B8 both bold and yellow background (0.25 pts)")
            total_score += 0.25
        else:
            issues = []
            if not b7_bold:
                issues.append(f"B7 not bold")
            if not b7_yellow:
                issues.append(f"B7 fill={repr(b7_fill)}, expected {YELLOW}")
            if not b8_bold:
                issues.append(f"B8 not bold")
            if not b8_yellow:
                issues.append(f"B8 fill={repr(b8_fill)}, expected {YELLOW}")
            print(f"FAIL: Component 3 — formatting issues: {'; '.join(issues)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: Quarterly hiring plan filled in rows 12-16 (0.25 pts)
    # Columns B-E (Q1-Q4) for rows 12-16 (region rows).
    # In the initial file, all these cells are None.
    # Golden has values summing to the ramp-adjusted rep count,
    # weighted toward H1 (Q1+Q2 >= Q3+Q4).
    # -----------------------------------------------------------------------
    try:
        q_values = []
        for row in range(12, 17):
            for col in range(2, 6):  # columns B(2) through E(5) = Q1-Q4
                val = ws.cell(row=row, column=col).value
                q_values.append(val)

        # All cells should be non-None integers/numbers
        non_null = [v for v in q_values if v is not None]

        if len(non_null) == 0:
            print(f"FAIL: Component 4 — quarterly hiring plan is entirely empty")
        else:
            # Check total is a reasonable positive number
            total_hires = sum(non_null)
            # Check H1 weighting: sum of Q1+Q2 should be >= sum of Q3+Q4
            q1_total = sum(ws.cell(row=r, column=2).value or 0 for r in range(12, 17))
            q2_total = sum(ws.cell(row=r, column=3).value or 0 for r in range(12, 17))
            q3_total = sum(ws.cell(row=r, column=4).value or 0 for r in range(12, 17))
            q4_total = sum(ws.cell(row=r, column=5).value or 0 for r in range(12, 17))

            h1_total = q1_total + q2_total
            h2_total = q3_total + q4_total

            all_filled = len(non_null) == len(q_values)
            h1_weighted = h1_total >= h2_total

            if total_hires > 0 and h1_weighted:
                print(f"PASS: Component 4 — quarterly plan filled ({len(non_null)}/20 cells), "
                      f"total hires={total_hires}, H1={h1_total} >= H2={h2_total} (0.25 pts)")
                total_score += 0.25
            elif total_hires > 0:
                print(f"FAIL: Component 4 — plan has values but not H1-weighted: "
                      f"H1={h1_total} < H2={h2_total}")
            else:
                print(f"FAIL: Component 4 — quarterly plan total is 0 or negative")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
