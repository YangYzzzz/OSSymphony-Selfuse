"""
Initial Setup: Sales Headcount and Capacity Planning Model
Task ID: calc_sales_headcount_planning_070
Domain: libreoffice_calc

Creates a spreadsheet with input parameters and quarterly hiring plan table,
but WITHOUT the calculated formulas (gap, reps needed) or filled-in hiring data.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_headcount_planning_070'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'HeadcountPlan'

    # ---- Section header ----
    ws['A1'] = 'Input Parameters'
    ws['A1'].font = Font(bold=True, size=12)

    # ---- Input parameters (A2:B8) ----
    # A2: Annual Revenue Target, B2: 24000000
    ws['A2'] = 'Annual Revenue Target'
    ws['B2'] = 24000000
    ws['B2'].number_format = '$#,##0'

    # A3: Current Total Quota Capacity, B3: 18000000
    ws['A3'] = 'Current Total Quota Capacity'
    ws['B3'] = 18000000
    ws['B3'].number_format = '$#,##0'

    # A4: Gap to Fill, B4: empty (agent will add formula)
    ws['A4'] = 'Gap to Fill'
    # B4 intentionally left empty

    # A5: Avg Quota per New Rep, B5: 600000
    ws['A5'] = 'Avg Quota per New Rep'
    ws['B5'] = 600000
    ws['B5'].number_format = '$#,##0'

    # A6: Ramp Factor (% of quota in year 1), B6: 0.7
    ws['A6'] = 'Ramp Factor (% of quota in year 1)'
    ws['B6'] = 0.7
    ws['B6'].number_format = '0%'

    # A7: New Reps Needed (fully ramped), B7: empty (agent will add formula)
    ws['A7'] = 'New Reps Needed (fully ramped)'
    # B7 intentionally left empty

    # A8: New Reps Needed (accounting for ramp), B8: empty (agent will add formula)
    ws['A8'] = 'New Reps Needed (accounting for ramp)'
    # B8 intentionally left empty

    # ---- Separator row A9 ----
    ws['A9'] = ''

    # ---- Quarterly Hiring Plan label ----
    ws['A10'] = 'Quarterly Hiring Plan'
    ws['A10'].font = Font(bold=True, size=12)

    # ---- Table headers (A11:E11) ----
    headers = ['Region', 'Q1', 'Q2', 'Q3', 'Q4']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=11, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        thin = Side(style='thin', color='000000')
        cell.border = Border(bottom=thin)

    # ---- Regional rows (12-16): region names only, Q1-Q4 empty ----
    regions = ['Northeast', 'Southeast', 'Midwest', 'West Coast', 'Southwest']
    for row_idx, region in enumerate(regions, 12):
        ws.cell(row=row_idx, column=1, value=region)
        # Q1-Q4 columns (B-E) left empty — agent fills in after computing B8

    # ---- Column widths ----
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
