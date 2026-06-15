"""
Initial Setup: Customer Feedback Excel + PDF Form Template
Task ID: osworld_multi_apps_excel_pdf_form_010
Domain: libreoffice_calc (multi-app: Calc + PDF)

Creates:
- /home/user/customer_feedback.xlsx: Customer feedback spreadsheet with survey data
- /home/user/Desktop/feedback_form_template.pdf: PDF form template with rating checkboxes
Opens customer_feedback.xlsx in LibreOffice Calc.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_excel_pdf_form_010'
OUTPUT_XLSX = f'{WORKDIR}/customer_feedback.xlsx'
DESKTOP = f'{WORKDIR}/Desktop'
TEMPLATE_PDF = f'{DESKTOP}/feedback_form_template.pdf'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_customer_feedback_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CustomerFeedback"

    # Header style
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2E75B6", end_color="FF2E75B6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["FeedbackID", "CustomerName", "ProductPurchased", "PurchaseDate",
               "OverallRating", "WouldRecommend", "Comments"]
    col_widths = [14, 22, 28, 14, 16, 16, 50]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
        ws.column_dimensions[chr(64 + col)].width = w

    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2"

    # Data: 12 realistic customer feedback rows
    data = [
        ("FB-2025-001", "Sarah Chen",       "UltraBoost Pro Headphones",   "2025-01-10", 5, "Y",
         "Absolutely love the sound quality! The noise cancellation is top-notch. Would highly recommend to anyone looking for premium audio."),
        ("FB-2025-002", "Marcus Johnson",   "SmartHome Hub X200",          "2025-01-15", 4, "Y",
         "Good product overall. Setup was straightforward and integration with other devices worked well. Minor issue with the app UI."),
        ("FB-2025-003", "Elena Rodriguez",  "ErgoDesk Standing Desk Pro",  "2025-01-18", 5, "Y",
         "This desk has transformed my workday! The motorized height adjustment is smooth and quiet. Very satisfied with the build quality."),
        ("FB-2025-004", "David Kim",        "FitTrack Smartwatch Series 3", "2025-01-22", 3, "N",
         "Battery life is disappointing - only lasts 1.5 days. Heart rate monitor is inaccurate during high-intensity workouts. Needs improvement."),
        ("FB-2025-005", "Priya Patel",      "CloudStor Portable SSD 2TB",  "2025-01-25", 5, "Y",
         "Incredibly fast transfer speeds! Compact design makes it perfect for travel. Survived a drop from desk height without any data loss."),
        ("FB-2025-006", "James O'Brien",    "AquaPure Water Filtration Sys","2025-02-02", 4, "Y",
         "Installation was easy and water tastes noticeably better. Filter replacement reminders via the app are a nice touch."),
        ("FB-2025-007", "Mei-Ling Zhang",   "LuminaDesk LED Monitor 27in",  "2025-02-05", 5, "Y",
         "Display colors are vibrant and accurate. The adjustable stand is very flexible. No dead pixels or backlight bleeding after 3 weeks of use."),
        ("FB-2025-008", "Robert Vasquez",   "SwiftKey Mechanical Keyboard", "2025-02-10", 4, "Y",
         "Tactile feedback is satisfying for typing. Build quality feels solid and premium. Slightly louder than expected but acceptable."),
        ("FB-2025-009", "Aisha Williams",   "NutriBlend Pro Blender 1200W", "2025-02-14", 2, "N",
         "Leaks around the blade assembly after 2 weeks. Customer service response was slow. The motor is powerful but the seal quality is poor."),
        ("FB-2025-010", "Thomas Nguyen",    "PowerCell Solar Charger 20W",  "2025-02-18", 4, "Y",
         "Works great for camping trips. Charges phone efficiently in direct sunlight. Foldable design is very convenient for backpacking."),
        ("FB-2025-011", "Sofia Kowalski",   "BrewMaster Coffee Station Pro","2025-02-22", 5, "Y",
         "Makes the best espresso I have ever had at home. The milk frother creates perfect microfoam. Worth every penny for coffee lovers."),
        ("FB-2025-012", "Nathan Brooks",    "CoolMax Air Purifier HEPA 500","2025-02-28", 3, "Y",
         "Air quality noticeably improved in my bedroom. A bit noisy on the highest setting but quiet on medium. App connectivity is unreliable."),
    ]

    data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal="center", vertical="top")

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = data_border
            if c in (5, 6):  # OverallRating, WouldRecommend - center
                cell.alignment = center_align
            elif c == 7:  # Comments - wrap text
                cell.alignment = data_alignment
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top")
        ws.row_dimensions[r].height = 40

    wb.save(OUTPUT_XLSX)
    print(f'Initial file created: {OUTPUT_XLSX}')


def create_pdf_template():
    """Create a PDF form template with star rating checkboxes (1-5) and Y/N recommend checkboxes."""
    try:
        from fpdf import FPDF

        os.makedirs(DESKTOP, exist_ok=True)

        pdf = FPDF()
        pdf.add_page()
        pdf.set_margins(20, 20, 20)

        # Title
        pdf.set_font("Helvetica", "B", 18)
        pdf.set_fill_color(46, 117, 182)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(0, 12, "Customer Feedback Survey Form", fill=True, align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(5)

        # Reset text color
        pdf.set_text_color(0, 0, 0)

        # Instructions
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(80, 80, 80)
        pdf.cell(0, 6, "Please complete this form based on your recent purchase experience.", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(5)

        pdf.set_text_color(0, 0, 0)

        # Section: Customer Information
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_fill_color(220, 235, 250)
        pdf.cell(0, 8, "Customer Information", fill=True, new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

        pdf.set_font("Helvetica", "", 11)
        # FeedbackID field
        pdf.cell(50, 8, "Feedback ID:")
        pdf.set_draw_color(0, 0, 0)
        pdf.cell(0, 8, "", border="B", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        # Customer Name field
        pdf.cell(50, 8, "Customer Name:")
        pdf.cell(0, 8, "", border="B", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        # Product Purchased field
        pdf.cell(50, 8, "Product Purchased:")
        pdf.cell(0, 8, "", border="B", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        # Purchase Date field
        pdf.cell(50, 8, "Purchase Date:")
        pdf.cell(0, 8, "", border="B", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(6)

        # Section: Overall Satisfaction Rating
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_fill_color(220, 235, 250)
        pdf.cell(0, 8, "Overall Satisfaction Rating", fill=True, new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

        pdf.set_font("Helvetica", "", 11)
        pdf.cell(0, 7, "Please mark your satisfaction level (1 = Very Dissatisfied, 5 = Very Satisfied):", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

        # Star rating checkboxes 1-5
        rating_labels = [
            ("1", "Very Dissatisfied"),
            ("2", "Dissatisfied"),
            ("3", "Neutral"),
            ("4", "Satisfied"),
            ("5", "Very Satisfied"),
        ]

        box_size = 6
        x_start = pdf.get_x()
        y_start = pdf.get_y()

        for rating, label in rating_labels:
            x = pdf.get_x()
            y = pdf.get_y()
            # Draw checkbox square
            pdf.rect(x, y + 1, box_size, box_size)
            pdf.set_x(x + box_size + 3)
            pdf.cell(0, 8, f"{rating} - {label}", new_x="LMARGIN", new_y="NEXT")
            pdf.ln(1)

        pdf.ln(4)

        # Section: Recommendation
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_fill_color(220, 235, 250)
        pdf.cell(0, 8, "Would You Recommend Us?", fill=True, new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

        pdf.set_font("Helvetica", "", 11)
        pdf.cell(0, 7, "Would you recommend our product/service to others?", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

        # Y/N checkboxes
        for option, text in [("Y", "Yes, I would recommend"), ("N", "No, I would not recommend")]:
            x = pdf.get_x()
            y = pdf.get_y()
            pdf.rect(x, y + 1, box_size, box_size)
            pdf.set_x(x + box_size + 3)
            pdf.cell(0, 8, f"{option} - {text}", new_x="LMARGIN", new_y="NEXT")
            pdf.ln(1)

        pdf.ln(4)

        # Section: Comments
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_fill_color(220, 235, 250)
        pdf.cell(0, 8, "Additional Comments", fill=True, new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

        pdf.set_font("Helvetica", "", 11)
        pdf.cell(0, 7, "Please share any additional feedback:", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

        # Comments text area (outlined box)
        pdf.set_draw_color(0, 0, 0)
        pdf.rect(pdf.get_x(), pdf.get_y(), 170, 40)
        pdf.ln(44)

        # Footer
        pdf.set_font("Helvetica", "I", 9)
        pdf.set_text_color(120, 120, 120)
        pdf.cell(0, 6, "Thank you for your feedback. Your response helps us improve our products and services.", align="C", new_x="LMARGIN", new_y="NEXT")

        pdf.output(TEMPLATE_PDF)
        print(f'PDF template created: {TEMPLATE_PDF}')

    except ImportError:
        print("fpdf2 not available, trying reportlab...")
        try:
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import mm

            os.makedirs(DESKTOP, exist_ok=True)
            c = canvas.Canvas(TEMPLATE_PDF, pagesize=A4)
            width, height = A4

            # Title
            c.setFont("Helvetica-Bold", 18)
            c.setFillColorRGB(0.18, 0.46, 0.71)
            c.drawCentredString(width / 2, height - 40 * mm, "Customer Feedback Survey Form")

            c.setFillColorRGB(0, 0, 0)
            c.setFont("Helvetica", 10)
            c.drawCentredString(width / 2, height - 50 * mm,
                                "Please complete this form based on your recent purchase experience.")

            y = height - 65 * mm

            # Customer Info section
            c.setFont("Helvetica-Bold", 12)
            c.setFillColorRGB(0.18, 0.46, 0.71)
            c.drawString(20 * mm, y, "Customer Information")
            y -= 10 * mm

            c.setFillColorRGB(0, 0, 0)
            c.setFont("Helvetica", 11)
            for label in ["Feedback ID:", "Customer Name:", "Product Purchased:", "Purchase Date:"]:
                c.drawString(20 * mm, y, label)
                c.line(70 * mm, y - 1, 190 * mm, y - 1)
                y -= 10 * mm

            y -= 5 * mm

            # Rating section
            c.setFont("Helvetica-Bold", 12)
            c.setFillColorRGB(0.18, 0.46, 0.71)
            c.drawString(20 * mm, y, "Overall Satisfaction Rating")
            y -= 10 * mm

            c.setFillColorRGB(0, 0, 0)
            c.setFont("Helvetica", 11)
            c.drawString(20 * mm, y, "Please mark your rating (1=Very Dissatisfied, 5=Very Satisfied):")
            y -= 10 * mm

            for rating, label in [("1", "Very Dissatisfied"), ("2", "Dissatisfied"),
                                   ("3", "Neutral"), ("4", "Satisfied"), ("5", "Very Satisfied")]:
                c.rect(20 * mm, y - 2, 5 * mm, 5 * mm)
                c.drawString(28 * mm, y, f"{rating} - {label}")
                y -= 9 * mm

            y -= 5 * mm

            # Recommendation section
            c.setFont("Helvetica-Bold", 12)
            c.setFillColorRGB(0.18, 0.46, 0.71)
            c.drawString(20 * mm, y, "Would You Recommend Us?")
            y -= 10 * mm

            c.setFillColorRGB(0, 0, 0)
            c.setFont("Helvetica", 11)
            c.drawString(20 * mm, y, "Would you recommend our product/service to others?")
            y -= 10 * mm

            for opt, text in [("Y", "Yes, I would recommend"), ("N", "No, I would not recommend")]:
                c.rect(20 * mm, y - 2, 5 * mm, 5 * mm)
                c.drawString(28 * mm, y, f"{opt} - {text}")
                y -= 9 * mm

            y -= 5 * mm

            # Comments section
            c.setFont("Helvetica-Bold", 12)
            c.setFillColorRGB(0.18, 0.46, 0.71)
            c.drawString(20 * mm, y, "Additional Comments")
            y -= 10 * mm

            c.setFillColorRGB(0, 0, 0)
            c.setFont("Helvetica", 11)
            c.drawString(20 * mm, y, "Please share any additional feedback:")
            y -= 10 * mm
            c.rect(20 * mm, y - 30 * mm, 170 * mm, 30 * mm)

            c.save()
            print(f'PDF template created: {TEMPLATE_PDF}')

        except ImportError:
            print("Neither fpdf2 nor reportlab available. Skipping PDF template creation.")


def main():
    # Create Desktop directory if needed
    os.makedirs(DESKTOP, exist_ok=True)

    # Create customer feedback Excel file
    create_customer_feedback_xlsx()

    # Create PDF form template on Desktop
    create_pdf_template()

    # GUI-ready startup: open customer_feedback.xlsx in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT_XLSX}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with customer_feedback.xlsx (DISPLAY=:0)')


main()
