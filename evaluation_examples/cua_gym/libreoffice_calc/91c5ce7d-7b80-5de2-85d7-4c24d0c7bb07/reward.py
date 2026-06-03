"""
Reward Script: Add year-over-year percentage change row to hospital budget table
Task ID: osworld_calc_annual_pct_change_004
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): Row 12 added with '% Change YoY' label in A12
  Component 2 (0.35): C12 and D12 contain percentage change formulas or numeric values
  Component 3 (0.30): Conditional formatting on C12:D12 with red fill for negative values
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_annual_pct_change_004'


def _check_pct_change_cell(cell_val, col_name):
    """
    Returns True if the cell value is a plausible percentage change calculation.
    Accepts:
      - Formula string starting with '=' that references appropriate columns and division
      - Numeric value in plausible range (-1.0 to 1.0 as decimal, or -100 to 100 as whole %)
    Returns False otherwise.
    """
    if cell_val is None:
        print(f"FAIL: {col_name} is empty (None)")
        return False

    if isinstance(cell_val, str) and cell_val.startswith('='):
        formula_upper = cell_val.upper().replace(' ', '')
        if '/' in formula_upper:
            print(f"PASS: {col_name} has percentage change formula: {repr(cell_val)}")
            return True
        print(f"FAIL: {col_name} formula lacks division operator: {repr(cell_val)}")
        return False

    if isinstance(cell_val, (int, float)):
        if -100.0 <= cell_val <= 100.0:
            print(f"PASS: {col_name} has numeric percentage value: {cell_val}")
            return True
        print(f"FAIL: {col_name} numeric value out of percentage range: {cell_val}")
        return False

    print(f"FAIL: {col_name} unexpected type {type(cell_val)}: {repr(cell_val)}")
    return False


def _is_negative_check_rule(rule):
    """
    Returns True if this CF rule checks for values less than 0.
    Checks both 'cellIs lessThan 0' and formula-based '<0' patterns.
    """
    rule_type = getattr(rule, 'type', None)
    rule_operator = getattr(rule, 'operator', None)
    rule_formula = getattr(rule, 'formula', None)

    if rule_type == 'cellIs' and rule_operator == 'lessThan':
        if rule_formula and len(rule_formula) > 0:
            try:
                return float(str(rule_formula[0])) == 0.0
            except (ValueError, TypeError):
                return False

    if rule_type in ('expression', 'formula') and rule_formula:
        formula_str = str(rule_formula).upper()
        return '<0' in formula_str or '< 0' in formula_str

    return False


def _has_red_fill_in_dxf(rule):
    """
    Returns True if the CF rule's dxf fill uses a red foreground color (contains FF0000).
    """
    dxf = getattr(rule, 'dxf', None)
    if not dxf:
        return False
    fill = getattr(dxf, 'fill', None)
    if not fill:
        return False
    try:
        fg_rgb = fill.fgColor.rgb
        return bool(fg_rgb and 'FF0000' in fg_rgb.upper())
    except Exception:
        return False


def _check_cf_red_on_negative(ws):
    """
    Check that conditional formatting exists covering row-12 percentage cells,
    applies when value < 0, and uses a red fill.
    Returns (cf_found, negative_check, red_fill) tuple of booleans.
    """
    cf_rules_map = ws.conditional_formatting._cf_rules

    for cf_range, cf_rules in cf_rules_map.items():
        cf_range_str = str(cf_range)
        # Must cover C12 or D12 in row 12
        if '12' not in cf_range_str:
            continue
        if 'C' not in cf_range_str and 'D' not in cf_range_str:
            continue

        for rule in cf_rules:
            if not _is_negative_check_rule(rule):
                continue
            # Rule is a negative-value check on the right cells
            has_red = _has_red_fill_in_dxf(rule)
            return (True, True, has_red)

    # Check if any CF at all on row 12 (but not a negative-value rule)
    for cf_range in cf_rules_map.keys():
        if '12' in str(cf_range):
            return (True, False, False)

    return (False, False, False)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get the worksheet
    try:
        if 'Hospital Budget' in wb.sheetnames:
            ws = wb['Hospital Budget']
        else:
            ws = wb.active
    except Exception as e:
        print(f"CRITICAL: Cannot access sheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: verify this is the correct file
    try:
        if ws['A1'].value != 'Department':
            print("CRITICAL: Unexpected file structure — A1 is not 'Department'")
            print("REWARD: 0.0")
            return 0.0
    except Exception as e:
        print(f"CRITICAL: Precondition check failed: {e}")
        print("REWARD: 0.0")
        return 0.0

    # -------------------------------------------------------------------------
    # Component 1: Row 12 added with '% Change YoY' label in A12 (0.35 points)
    # This FAILS on initial (row 12 is empty) and PASSES on golden (A12 has label)
    # -------------------------------------------------------------------------
    try:
        a12_value = ws.cell(row=12, column=1).value
        if a12_value is not None and '% change' in str(a12_value).lower():
            print(f"PASS: Component 1 — A12 has '% Change YoY' label: {repr(a12_value)} (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 1 — A12 expected '% Change YoY' label, found: {repr(a12_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: C12 and D12 contain percentage change calculations (0.35 points)
    # Task requires year-over-year percentage change formulas or computed values.
    # This FAILS on initial (C12 and D12 are empty) and PASSES on golden.
    # -------------------------------------------------------------------------
    try:
        c12_val = ws.cell(row=12, column=3).value
        d12_val = ws.cell(row=12, column=4).value

        c12_ok = _check_pct_change_cell(c12_val, 'C12')
        d12_ok = _check_pct_change_cell(d12_val, 'D12')

        if c12_ok and d12_ok:
            print(f"PASS: Component 2 — Both C12 and D12 have valid percentage change values (0.35 pts)")
            total_score += 0.35
        elif c12_ok or d12_ok:
            print(f"PARTIAL: Component 2 — Only one pct change cell valid (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 — Neither C12 nor D12 has valid percentage change values")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Conditional formatting with red fill for negative % changes (0.30 points)
    # Task requires negative percentage cells highlighted in red.
    # This FAILS on initial (no CF exists) and PASSES on golden.
    # -------------------------------------------------------------------------
    try:
        cf_found, negative_check, red_fill = _check_cf_red_on_negative(ws)

        if cf_found and negative_check and red_fill:
            print(f"PASS: Component 3 — CF rule for negative values with red fill present (0.30 pts)")
            total_score += 0.30
        elif cf_found and negative_check:
            print(f"PARTIAL: Component 3 — CF for negative values found but red fill not confirmed (0.15 pts)")
            total_score += 0.15
        elif cf_found:
            print(f"PARTIAL: Component 3 — CF on row 12 found but not checking for negatives (0.10 pts)")
            total_score += 0.10
        else:
            cf_ranges = list(ws.conditional_formatting._cf_rules.keys())
            if cf_ranges:
                print(f"FAIL: Component 3 — CF exists but not on row 12 pct cells (ranges: {[str(r) for r in cf_ranges]})")
            else:
                print(f"FAIL: Component 3 — No conditional formatting found anywhere in sheet")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Main entry point — test against canonical artifact path on VM
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
