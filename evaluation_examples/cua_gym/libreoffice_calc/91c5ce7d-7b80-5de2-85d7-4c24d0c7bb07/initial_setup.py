"""
Initial Setup: Add year-over-year percentage change row to hospital budget table
Task ID: osworld_calc_annual_pct_change_004
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_annual_pct_change_004'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Hospital Budget ---
    ws = wb.active
    ws.title = 'Hospital Budget'

    # Header row styling
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font = Font(name='Calibri', bold=True, color='FFFFFFFF', size=11)
    center_align = Alignment(horizontal='center', vertical='center')

    # Column headers
    headers = ['Department', '2022 Budget ($)', '2023 Budget ($)', '2024 Budget ($)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Realistic hospital department budget data (amounts in USD)
    departments = [
        ('Emergency Medicine',   4850000, 5120000, 5380000),
        ('Cardiology',           6230000, 6580000, 6210000),  # decline in 2024
        ('Radiology',            3970000, 4150000, 4430000),
        ('Oncology',             7410000, 7820000, 8290000),
        ('Pediatrics',           3120000, 3350000, 3180000),  # decline in 2024
        ('Orthopedics',          4560000, 4740000, 4920000),
        ('Neurology',            5890000, 6150000, 6470000),
        ('General Surgery',      5140000, 5380000, 5020000),  # decline in 2024
        ('Psychiatry',           2780000, 2940000, 3110000),
        ('Pharmacy',             3650000, 3890000, 4060000),
    ]

    dept_font = Font(name='Calibri', size=11)
    num_font = Font(name='Calibri', size=11)
    dept_align = Alignment(horizontal='left', vertical='center')
    num_align = Alignment(horizontal='right', vertical='center')

    # Alternate row fill
    light_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    no_fill = PatternFill(fill_type=None)

    for r, (dept, b2022, b2023, b2024) in enumerate(departments, 2):
        row_fill = light_fill if r % 2 == 0 else no_fill

        dept_cell = ws.cell(row=r, column=1, value=dept)
        dept_cell.font = dept_font
        dept_cell.alignment = dept_align
        dept_cell.fill = row_fill

        for col, val in enumerate([b2022, b2023, b2024], 2):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = num_font
            cell.number_format = '$#,##0'
            cell.alignment = num_align
            cell.fill = row_fill

    # NOTE: NO "% Change YoY" row — that is the task to add
    # NOTE: NO conditional formatting — that is the task to add

    # Column widths
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    # Row height for header
    ws.row_dimensions[1].height = 20

    # Freeze the header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup — open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
