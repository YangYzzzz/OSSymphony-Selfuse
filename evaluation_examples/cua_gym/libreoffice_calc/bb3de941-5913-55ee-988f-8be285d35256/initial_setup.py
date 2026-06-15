"""
Initial Setup: Inventory turnover analysis with COGS data across sheets
Task ID: calc_ops_inventory_turnover_045
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_inventory_turnover_045'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: COGSData ---
    ws_cogs = wb.active
    ws_cogs.title = 'COGSData'

    # Headers
    cogs_headers = ['Category', 'Q1 COGS', 'Q2 COGS', 'Q3 COGS', 'Q4 COGS', 'Annual COGS']
    for col, h in enumerate(cogs_headers, 1):
        cell = ws_cogs.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # 5 product categories with quarterly COGS data (Annual COGS column F is intentionally empty)
    cogs_data = [
        ['Electronics',    320500, 298700, 415200, 502300],
        ['Apparel',        148200, 162300, 189500, 210400],
        ['Home & Garden',   95400, 102800,  88600, 115300],
        ['Sporting Goods',  74300,  81200,  93700,  68900],
        ['Food & Beverage', 521000, 489300, 563400, 610200],
    ]
    for r, row_data in enumerate(cogs_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_cogs.cell(row=r, column=c, value=val)
        # Column F (Annual COGS) intentionally left empty

    ws_cogs.column_dimensions['A'].width = 22
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws_cogs.column_dimensions[col_letter].width = 15

    # --- Sheet 2: AvgInventory ---
    ws_inv = wb.create_sheet('AvgInventory')

    # Headers
    inv_headers = ['Category', 'Avg Inventory Value']
    for col, h in enumerate(inv_headers, 1):
        cell = ws_inv.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Matching categories with average inventory values
    inv_data = [
        ['Electronics',     185000],
        ['Apparel',          68500],
        ['Home & Garden',    42300],
        ['Sporting Goods',   55800],
        ['Food & Beverage',  97400],
    ]
    for r, row_data in enumerate(inv_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_inv.cell(row=r, column=c, value=val)

    ws_inv.column_dimensions['A'].width = 22
    ws_inv.column_dimensions['B'].width = 22

    # --- Sheet 3: TurnoverAnalysis ---
    ws_turn = wb.create_sheet('TurnoverAnalysis')

    # Headers
    turn_headers = [
        'Category', 'Annual COGS', 'Avg Inventory',
        'Turnover Ratio', 'Days Inventory Outstanding', 'Below Target'
    ]
    for col, h in enumerate(turn_headers, 1):
        cell = ws_turn.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Category names in column A only — B through F intentionally empty
    categories = ['Electronics', 'Apparel', 'Home & Garden', 'Sporting Goods', 'Food & Beverage']
    for r, cat in enumerate(categories, 2):
        ws_turn.cell(row=r, column=1, value=cat)

    ws_turn.column_dimensions['A'].width = 22
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws_turn.column_dimensions[col_letter].width = 25

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: COGSData (Q1-Q4 COGS, Annual COGS empty), AvgInventory, TurnoverAnalysis (only categories filled)')


create_initial()
