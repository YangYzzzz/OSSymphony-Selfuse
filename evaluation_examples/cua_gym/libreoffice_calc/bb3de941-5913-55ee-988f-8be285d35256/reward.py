"""
Reward Script: Calculate inventory turnover ratio and days inventory outstanding
Task ID: calc_ops_inventory_turnover_045
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): COGSData F2:F6 have Annual COGS sum formulas (=B+C+D+E)
  Component 2 (0.25): TurnoverAnalysis B2:B6 VLOOKUP to COGSData, C2:C6 VLOOKUP to AvgInventory
  Component 3 (0.25): TurnoverAnalysis D2:D6 turnover ratio (=B/C) and E2:E6 days outstanding (=365/D)
  Component 4 (0.25): TurnoverAnalysis F2:F6 IF formula for below-target flag + conditional formatting
"""

import os
import openpyxl
import re

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_ops_inventory_turnover_045'


def normalize_formula(f):
    """Normalize formula for flexible comparison: remove spaces, uppercase."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def formula_matches_pattern(cell_value, pattern_re):
    """Check if cell formula matches a regex pattern (case-insensitive, spaces removed)."""
    if not isinstance(cell_value, str):
        return False
    normalized = normalize_formula(cell_value)
    return bool(re.match(pattern_re, normalized, re.IGNORECASE))


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: required sheets exist
    required_sheets = ['COGSData', 'AvgInventory', 'TurnoverAnalysis']
    for sheet_name in required_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"CRITICAL: Required sheet '{sheet_name}' not found. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0

    ws_cogs = wb['COGSData']
    ws_ta = wb['TurnoverAnalysis']

    # -------------------------------------------------------------------------
    # Component 1: COGSData F2:F6 Annual COGS sum formulas (0.25 points)
    # Each row should have =B+C+D+E (sum of quarterly COGS).
    # Pattern: =B{row}+C{row}+D{row}+E{row}  (order may vary, but all 4 quarters)
    # -------------------------------------------------------------------------
    try:
        cogs_formula_count = 0
        for row in range(2, 7):
            cell_val = ws_cogs.cell(row=row, column=6).value  # Column F
            if isinstance(cell_val, str) and cell_val.startswith('='):
                # Verify it sums the four quarterly columns B, C, D, E
                norm = normalize_formula(cell_val)
                # Accept patterns like =B2+C2+D2+E2 or =SUM(B2:E2) etc.
                # Require all 4 columns B,C,D,E to be referenced
                has_b = f'B{row}' in norm
                has_c = f'C{row}' in norm
                has_d = f'D{row}' in norm
                has_e = f'E{row}' in norm
                # Also accept SUM(B{row}:E{row})
                has_sum_range = f'B{row}:E{row}' in norm or f'SUM(B{row}:E{row})' in norm
                if (has_b and has_c and has_d and has_e) or has_sum_range:
                    cogs_formula_count += 1
                else:
                    print(f"FAIL: COGSData F{row} formula '{cell_val}' does not sum B+C+D+E")
            else:
                print(f"FAIL: COGSData F{row} is not a formula: {repr(cell_val)}")

        if cogs_formula_count == 5:
            print(f"PASS: Component 1 — COGSData F2:F6 all have Annual COGS sum formulas (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — COGSData F2:F6: {cogs_formula_count}/5 have correct sum formulas")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: TurnoverAnalysis B2:B6 VLOOKUP to COGSData Annual COGS,
    #              C2:C6 VLOOKUP to AvgInventory (0.25 points)
    # -------------------------------------------------------------------------
    try:
        vlookup_b_count = 0
        vlookup_c_count = 0
        for row in range(2, 7):
            # B column: VLOOKUP into COGSData for annual COGS (column 6 of COGSData)
            b_val = ws_ta.cell(row=row, column=2).value
            if isinstance(b_val, str) and 'VLOOKUP' in b_val.upper():
                norm = normalize_formula(b_val)
                # Should reference COGSData and column index 6
                if 'COGSDATA' in norm and '6' in norm:
                    vlookup_b_count += 1
                else:
                    print(f"FAIL: TurnoverAnalysis B{row} VLOOKUP missing COGSData or col 6: {repr(b_val)}")
            else:
                print(f"FAIL: TurnoverAnalysis B{row} is not a VLOOKUP: {repr(b_val)}")

            # C column: VLOOKUP into AvgInventory for avg inventory value (column 2)
            c_val = ws_ta.cell(row=row, column=3).value
            if isinstance(c_val, str) and 'VLOOKUP' in c_val.upper():
                norm = normalize_formula(c_val)
                if 'AVGINVENTORY' in norm and '2' in norm:
                    vlookup_c_count += 1
                else:
                    print(f"FAIL: TurnoverAnalysis C{row} VLOOKUP missing AvgInventory or col 2: {repr(c_val)}")
            else:
                print(f"FAIL: TurnoverAnalysis C{row} is not a VLOOKUP: {repr(c_val)}")

        if vlookup_b_count == 5 and vlookup_c_count == 5:
            print(f"PASS: Component 2 — TurnoverAnalysis B2:B6 and C2:C6 all have correct VLOOKUPs (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — B VLOOKUPs: {vlookup_b_count}/5, C VLOOKUPs: {vlookup_c_count}/5")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: TurnoverAnalysis D2:D6 = B/C (turnover ratio),
    #              E2:E6 = 365/D (days inventory outstanding) (0.25 points)
    # -------------------------------------------------------------------------
    try:
        ratio_count = 0
        dio_count = 0
        for row in range(2, 7):
            # D column: =B{row}/C{row} (turnover ratio)
            d_val = ws_ta.cell(row=row, column=4).value
            if isinstance(d_val, str) and d_val.startswith('='):
                norm = normalize_formula(d_val)
                # Accept =B{row}/C{row}
                expected_ratio = f'=B{row}/C{row}'
                if normalize_formula(expected_ratio) == norm:
                    ratio_count += 1
                else:
                    print(f"FAIL: TurnoverAnalysis D{row} formula '{d_val}' does not match =B{row}/C{row}")
            else:
                print(f"FAIL: TurnoverAnalysis D{row} is not a formula: {repr(d_val)}")

            # E column: =365/D{row} (days inventory outstanding)
            e_val = ws_ta.cell(row=row, column=5).value
            if isinstance(e_val, str) and e_val.startswith('='):
                norm = normalize_formula(e_val)
                expected_dio = f'=365/D{row}'
                if normalize_formula(expected_dio) == norm:
                    dio_count += 1
                else:
                    print(f"FAIL: TurnoverAnalysis E{row} formula '{e_val}' does not match =365/D{row}")
            else:
                print(f"FAIL: TurnoverAnalysis E{row} is not a formula: {repr(e_val)}")

        if ratio_count == 5 and dio_count == 5:
            print(f"PASS: Component 3 — TurnoverAnalysis D2:D6 (turnover ratio) and E2:E6 (DIO) correct (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 3 — Turnover ratio: {ratio_count}/5, DIO: {dio_count}/5")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: TurnoverAnalysis F2:F6 IF formula for below-target flag
    #              AND conditional formatting with red fill for 'BELOW TARGET' (0.25 points)
    # -------------------------------------------------------------------------
    try:
        if_formula_count = 0
        for row in range(2, 7):
            f_val = ws_ta.cell(row=row, column=6).value
            if isinstance(f_val, str) and 'IF' in f_val.upper():
                norm = normalize_formula(f_val)
                # Must check D<6 and produce "BELOW TARGET" vs "OK"
                has_d_lt_6 = f'D{row}<6' in norm
                has_below_target = 'BELOWTARGET' in norm.replace('"', '').replace("'", '')
                has_ok = 'OK' in norm
                if has_d_lt_6 and has_below_target and has_ok:
                    if_formula_count += 1
                else:
                    print(f"FAIL: TurnoverAnalysis F{row} IF formula missing expected logic: {repr(f_val)}")
            else:
                print(f"FAIL: TurnoverAnalysis F{row} is not an IF formula: {repr(f_val)}")

        # Check conditional formatting: red fill on F column for "BELOW TARGET"
        cf_found = False
        cf_has_red = False
        try:
            for cf_range in ws_ta.conditional_formatting:
                cf_range_str = str(cf_range)
                # Check if it covers column F (F2:F6 or F2:F7 etc.)
                if 'F' in cf_range_str:
                    for rule in ws_ta.conditional_formatting[cf_range]:
                        if rule.type in ('expression', 'formula'):
                            formula_str = ' '.join(rule.formula) if rule.formula else ''
                            if 'BELOW TARGET' in formula_str.upper() or 'BELOWTARGET' in formula_str.upper().replace('"', ''):
                                cf_found = True
                                # Check for red fill
                                if hasattr(rule, 'dxf') and rule.dxf and hasattr(rule.dxf, 'fill') and rule.dxf.fill:
                                    try:
                                        fg_rgb = rule.dxf.fill.fgColor.rgb
                                        # Red: FFFF0000 or any red-dominant color
                                        if fg_rgb and fg_rgb.upper().endswith('FF0000'):
                                            cf_has_red = True
                                    except Exception:
                                        pass
        except Exception as cf_err:
            print(f"  WARNING: Could not check conditional formatting: {cf_err}")

        if if_formula_count == 5 and cf_found and cf_has_red:
            print(f"PASS: Component 4 — F2:F6 IF formulas correct + red conditional formatting (0.25 pts)")
            total_score += 0.25
        elif if_formula_count == 5 and cf_found:
            print(f"PASS (partial): Component 4 — F2:F6 IF formulas correct + conditional formatting exists, but fill color not confirmed red (0.20 pts)")
            total_score += 0.20
        elif if_formula_count == 5:
            print(f"PARTIAL: Component 4 — F2:F6 IF formulas correct (0.15 pts), but conditional formatting missing")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — IF formulas: {if_formula_count}/5, CF found: {cf_found}, CF red: {cf_has_red}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
