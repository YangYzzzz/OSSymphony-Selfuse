"""
Initial Setup: Auto-fit row heights task
Task ID: calc_fmt_row_autofit_050
Domain: libreoffice_calc

Creates a spreadsheet 'Notes Database' with 29 rows of notes data (rows 2-30).
All rows 2-30 are set to a fixed height of 15pt (simulating clipped/excessive whitespace).
Column C has text wrap enabled with multi-line text content.
"""

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_row_autofit_050'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Notes Database'

    # --- Headers (Row 1) ---
    headers = ['ID', 'Category', 'Notes', 'Date']
    header_font = Font(bold=True, name='Calibri', size=11)
    header_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

    # Row 1 height stays at default (will remain unchanged after task)
    ws.row_dimensions[1].height = 20

    # --- Data Rows (2-30) ---
    # Multi-line notes data with varying lengths to demonstrate auto-fit need
    data = [
        (1,  'Bug Report',       'Application crashes when user clicks the Export button twice in quick succession.\nThis only happens on Windows 10 and above.\nWorkaround: wait 2 seconds between clicks.',                                                                              '2025-01-05'),
        (2,  'Feature Request',  'Add dark mode support to the main dashboard.\nUsers have requested this multiple times in feedback surveys.\nEstimated effort: 2 sprints.',                                                                                                            '2025-01-08'),
        (3,  'Meeting Notes',    'Q1 planning session: agreed on 3 major deliverables.\nTimeline pushed back by 2 weeks due to resource constraints.\nAction items assigned to Sarah Chen and Marcus Johnson.',                                                                          '2025-01-10'),
        (4,  'Task',             'Update API documentation.',                                                                                                                                                                                                                           '2025-01-12'),
        (5,  'Bug Report',       'Login page fails to load when cookies are disabled in browser settings.\nReproducible on Chrome 120+ and Firefox 121.\nHigh priority – affects 5% of user base.\nAssigned to backend team for investigation.',                                        '2025-01-15'),
        (6,  'Research',         'Evaluated three third-party charting libraries: Chart.js, D3.js, and ECharts.\nChart.js is simplest to integrate but limited in 3D support.\nRecommendation: use ECharts for new dashboard module.',                                                  '2025-01-17'),
        (7,  'Feature Request',  'Export reports to PDF format directly from the application.',                                                                                                                                                                                         '2025-01-20'),
        (8,  'Bug Report',       'Date picker component does not respect locale settings.\nUS users see MM/DD/YYYY, EU users see MM/DD/YYYY instead of DD/MM/YYYY.\nFix applied in dev branch, pending QA review.',                                                                     '2025-01-22'),
        (9,  'Meeting Notes',    'Sprint retrospective highlights:\n- Velocity improved by 15% compared to last sprint\n- Team morale is high after successful product demo\n- Two blockers identified: slow CI pipeline and unclear requirements for module B',                         '2025-01-25'),
        (10, 'Task',             'Migrate legacy authentication service to OAuth 2.0.\nInvolves updating 12 microservices.\nEstimated completion: end of Q2.',                                                                                                                          '2025-01-28'),
        (11, 'Research',         'Conducted user interviews with 10 participants.\nKey insights:\n1. Users want faster search\n2. Navigation is confusing for new users\n3. Mobile experience needs major improvement\n4. Notification settings are hard to find.',                     '2025-02-01'),
        (12, 'Bug Report',       'Memory leak detected in the image processing module.',                                                                                                                                                                                                '2025-02-03'),
        (13, 'Feature Request',  'Allow users to customize dashboard widgets by drag-and-drop.\nThis would greatly improve personalization and workflow efficiency.\nComparable feature exists in competitor product.\nEstimated effort: 4 sprints including testing.',                   '2025-02-05'),
        (14, 'Task',             'Set up automated performance benchmarking for nightly builds.',                                                                                                                                                                                       '2025-02-07'),
        (15, 'Meeting Notes',    'Product roadmap review:\n- Q2: Mobile app v2.0 release\n- Q3: Enterprise SSO integration\n- Q4: AI-powered search feature\nBudget approved for all three initiatives.',                                                                              '2025-02-10'),
        (16, 'Bug Report',       'Notification emails are being sent with incorrect timestamps.\nThe server is using UTC but displaying as local time without conversion.\nAffects all users in non-UTC timezones.\nPatch in progress.',                                                 '2025-02-12'),
        (17, 'Research',         'Security audit findings: 3 medium-severity vulnerabilities identified in the authentication flow.\nAll issues related to session management.\nRecommended fixes documented in security report v2.3.',                                                  '2025-02-14'),
        (18, 'Task',             'Write unit tests for the payment processing module.',                                                                                                                                                                                                 '2025-02-17'),
        (19, 'Feature Request',  'Add multi-language support (i18n) for at least 5 languages: Spanish, French, German, Japanese, and Portuguese.\nThis is critical for expanding into new markets.\nNeeds translation service integration.',                                             '2025-02-19'),
        (20, 'Bug Report',       'CSV import fails silently when file contains special characters (e.g., accented letters, Chinese characters).\nNo error message shown to user.\nData is partially imported without warning.',                                                          '2025-02-21'),
        (21, 'Meeting Notes',    'Design review for new onboarding flow:\n- Reduce steps from 7 to 4\n- Add interactive tutorial\n- Show progress indicator throughout\n- Allow users to skip optional sections\nApproved for development.',                                           '2025-02-24'),
        (22, 'Research',         'Benchmarked database query performance after index optimization.\nRead queries improved by 40%.\nWrite operations slightly slower due to index maintenance.\nOverall recommendation: proceed with indexing strategy.',                                  '2025-02-26'),
        (23, 'Task',             'Update all third-party dependencies to latest stable versions.',                                                                                                                                                                                      '2025-03-01'),
        (24, 'Bug Report',       'Search results do not update when filters are changed without pressing the Search button.\nExpected behavior: auto-refresh on filter change.\nThis is a UX regression from version 3.1.',                                                             '2025-03-03'),
        (25, 'Feature Request',  'Implement real-time collaboration features allowing multiple users to edit the same document simultaneously.',                                                                                                                                         '2025-03-05'),
        (26, 'Research',         'Analyzed customer support tickets from January to March.\n287 tickets total.\nTop issues:\n1. Password reset (22%)\n2. Billing questions (18%)\n3. Export functionality (15%)\n4. Performance complaints (14%)\nSuggested FAQ improvements attached.',  '2025-03-07'),
        (27, 'Meeting Notes',    'Cross-team sync: aligned on API versioning strategy.\nDecision: maintain v1 and v2 APIs in parallel for 6 months.\nDeprecation notices to be sent to all external developers by end of month.',                                                       '2025-03-10'),
        (28, 'Task',             'Deploy staging environment updates and run regression test suite.',                                                                                                                                                                                   '2025-03-12'),
        (29, 'Bug Report',       'User profile pictures are not resized correctly on mobile devices.\nImages appear stretched or pixelated on screens with high DPI.\nNeeds CSS fix and server-side image optimization.',                                                                '2025-03-14'),
    ]

    for row_data in data:
        row_num = row_data[0] + 1  # data rows start at 2
        ws.cell(row=row_num, column=1, value=row_data[0])  # ID
        ws.cell(row=row_num, column=2, value=row_data[1])  # Category
        # Column C: Notes with wrap_text=True
        notes_cell = ws.cell(row=row_num, column=3, value=row_data[2])
        notes_cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(row=row_num, column=4, value=row_data[3])  # Date

        # Fixed height of 15pt for all data rows (causing clipped / excess whitespace)
        ws.row_dimensions[row_num].height = 15

    # --- Column widths ---
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
