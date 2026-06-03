"""
Initial Setup: Apply conditional formatting for OVERDUE invoices
Task ID: calc_gfl_056
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_056'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Open Invoices"

    # Headers
    headers = ["Invoice No", "Customer", "Issue Date", "Due Date", "Amount", "Paid", "Status"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 34 invoice records (rows 2-35)
    # About 8 rows will have 'OVERDUE' status
    invoices = [
        ["INV-2025-001", "Meridian Healthcare Group", "2025-01-05", "2025-02-04", 12450.00, 12450.00, "Paid"],
        ["INV-2025-002", "Apex Manufacturing Ltd", "2025-01-08", "2025-02-07", 8730.50, 0.00, "OVERDUE"],
        ["INV-2025-003", "Silverlake Consulting", "2025-01-12", "2025-02-11", 3200.00, 3200.00, "Paid"],
        ["INV-2025-004", "Northwind Traders Inc", "2025-01-15", "2025-02-14", 15680.00, 0.00, "Pending"],
        ["INV-2025-005", "Coastal Dynamics Corp", "2025-01-18", "2025-03-19", 6425.75, 0.00, "Draft"],
        ["INV-2025-006", "Summit Financial Services", "2025-01-20", "2025-02-19", 22100.00, 0.00, "OVERDUE"],
        ["INV-2025-007", "BluePeak Technologies", "2025-01-22", "2025-02-21", 4890.00, 4890.00, "Paid"],
        ["INV-2025-008", "Ironclad Security Solutions", "2025-01-25", "2025-02-24", 9350.00, 9350.00, "Paid"],
        ["INV-2025-009", "GreenValley Organics", "2025-01-28", "2025-02-27", 1875.25, 0.00, "OVERDUE"],
        ["INV-2025-010", "Horizon Media Partners", "2025-02-01", "2025-03-03", 31200.00, 31200.00, "Paid"],
        ["INV-2025-011", "Cascade Logistics LLC", "2025-02-03", "2025-03-05", 7640.00, 0.00, "Pending"],
        ["INV-2025-012", "Vertex Engineering", "2025-02-05", "2025-03-07", 18900.50, 0.00, "OVERDUE"],
        ["INV-2025-013", "Starline Hospitality", "2025-02-07", "2025-03-09", 5430.00, 5430.00, "Paid"],
        ["INV-2025-014", "Redwood Analytics", "2025-02-10", "2025-03-12", 11275.00, 0.00, "Pending"],
        ["INV-2025-015", "Pacific Rim Exports", "2025-02-12", "2025-03-14", 42500.00, 42500.00, "Paid"],
        ["INV-2025-016", "Quantum Data Systems", "2025-02-14", "2025-03-16", 6780.25, 0.00, "OVERDUE"],
        ["INV-2025-017", "Atlas Construction Group", "2025-02-16", "2025-03-18", 29400.00, 0.00, "Draft"],
        ["INV-2025-018", "Evergreen Properties", "2025-02-18", "2025-03-20", 8125.00, 8125.00, "Paid"],
        ["INV-2025-019", "Titanium Aerospace Inc", "2025-02-20", "2025-03-22", 54300.00, 0.00, "Pending"],
        ["INV-2025-020", "Sapphire Healthcare", "2025-02-22", "2025-03-24", 3960.75, 0.00, "OVERDUE"],
        ["INV-2025-021", "Cobalt Mining Corp", "2025-02-24", "2025-03-26", 17850.00, 17850.00, "Paid"],
        ["INV-2025-022", "Pinnacle Insurance", "2025-02-26", "2025-03-28", 6200.00, 0.00, "Pending"],
        ["INV-2025-023", "Oakwood Furniture Co", "2025-02-28", "2025-03-30", 2340.50, 0.00, "Draft"],
        ["INV-2025-024", "Sterling Pharmaceuticals", "2025-03-02", "2025-04-01", 38750.00, 0.00, "OVERDUE"],
        ["INV-2025-025", "Crescent Bay Marina", "2025-03-04", "2025-04-03", 5125.00, 5125.00, "Paid"],
        ["INV-2025-026", "Iron Bridge Fabrication", "2025-03-06", "2025-04-05", 14600.00, 0.00, "Pending"],
        ["INV-2025-027", "Wildflower Catering", "2025-03-08", "2025-04-07", 3780.00, 3780.00, "Paid"],
        ["INV-2025-028", "Nexus Telecom Solutions", "2025-03-10", "2025-04-09", 21350.75, 0.00, "OVERDUE"],
        ["INV-2025-029", "Granite Peak Adventures", "2025-03-12", "2025-04-11", 8900.00, 0.00, "Pending"],
        ["INV-2025-030", "Lakeshore Dental Clinic", "2025-03-14", "2025-04-13", 1650.00, 1650.00, "Paid"],
        ["INV-2025-031", "Crimson Fox Creative", "2025-03-16", "2025-04-15", 7420.00, 0.00, "Draft"],
        ["INV-2025-032", "Brightstar Solar Energy", "2025-03-18", "2025-04-17", 46200.00, 46200.00, "Paid"],
        ["INV-2025-033", "Trident Marine Services", "2025-03-20", "2025-04-19", 9875.50, 0.00, "Pending"],
        ["INV-2025-034", "Canyon Ridge Vineyards", "2025-03-22", "2025-04-21", 4530.00, 4530.00, "Paid"],
    ]

    for r, row_data in enumerate(invoices, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 12

    # NO conditional formatting -- that is the task

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
