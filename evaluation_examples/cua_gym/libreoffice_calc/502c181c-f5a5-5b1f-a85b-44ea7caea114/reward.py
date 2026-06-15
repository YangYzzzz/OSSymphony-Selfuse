"""
Reward Script: Apply bold + dark red conditional formatting for OVERDUE rows
Task ID: calc_gfl_056
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Conditional formatting rule exists on the sheet
  Component 2 (0.25): Formula references column G for "OVERDUE"
  Component 3 (0.15): Rule range covers data area A2:G35
  Component 4 (0.15): DXF font is bold
  Component 5 (0.15): DXF font color is dark red
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_056'


def persist_app_state(domain: str):
    """Save any unsaved LibreOffice state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify that conditional formatting has been applied for OVERDUE rows.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get the sheet
    try:
        ws = wb['Open Invoices']
    except KeyError:
        print("CRITICAL: Sheet 'Open Invoices' not found")
        print("REWARD: 0.0")
        return 0.0

    # Gather all conditional formatting rules
    cf_list = list(ws.conditional_formatting)

    # Component 1: Conditional formatting rule exists (0.30 points)
    # This should FAIL on initial (no CF rules) and PASS on golden (has CF rule)
    try:
        if len(cf_list) > 0:
            # Further check: at least one rule must be expression/formula type
            has_expression_rule = False
            for cf in cf_list:
                for rule in cf.rules:
                    if rule.type in ('expression', 'formula'):
                        has_expression_rule = True
                        break
                if has_expression_rule:
                    break
            if has_expression_rule:
                print(f"PASS: Component 1 -- Found {len(cf_list)} conditional formatting entry/entries with expression rule (0.30 pts)")
                total_score += 0.30
            else:
                print(f"FAIL: Component 1 -- Found CF rules but none are expression/formula type")
        else:
            print("FAIL: Component 1 -- No conditional formatting rules found")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Formula references column G and checks for "OVERDUE" (0.25 points)
    try:
        formula_correct = False
        for cf in cf_list:
            for rule in cf.rules:
                if rule.type in ('expression', 'formula') and rule.formula:
                    for f in rule.formula:
                        f_clean = f.upper().replace(' ', '')
                        # Check formula references $G (column G) and "OVERDUE"
                        if '$G' in f_clean and 'OVERDUE' in f_clean:
                            formula_correct = True
                            print(f"PASS: Component 2 -- Formula '{f}' correctly references column G for OVERDUE (0.25 pts)")
                            break
                if formula_correct:
                    break
            if formula_correct:
                break
        if not formula_correct:
            # Also accept formulas without $ anchor but still referencing G column
            for cf in cf_list:
                for rule in cf.rules:
                    if rule.type in ('expression', 'formula') and rule.formula:
                        for f in rule.formula:
                            f_clean = f.upper().replace(' ', '')
                            if re.search(r'G\d', f_clean) and 'OVERDUE' in f_clean:
                                formula_correct = True
                                print(f"PASS: Component 2 -- Formula '{f}' references column G for OVERDUE (0.25 pts)")
                                break
                    if formula_correct:
                        break
                if formula_correct:
                    break
        if formula_correct:
            total_score += 0.25
        else:
            all_formulas = []
            for cf in cf_list:
                for rule in cf.rules:
                    if rule.formula:
                        all_formulas.extend(rule.formula)
            print(f"FAIL: Component 2 -- No formula referencing column G and OVERDUE found. Found formulas: {all_formulas}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Rule range covers data area (0.15 points)
    # Should cover at least A2:G35 or equivalent
    try:
        range_ok = False
        for cf in cf_list:
            has_expression = any(r.type in ('expression', 'formula') for r in cf.rules)
            if not has_expression:
                continue
            # Use cf.cells.ranges to get proper CellRange objects
            for rng in cf.cells.ranges:
                if rng.min_col == 1 and rng.min_row <= 2 and rng.max_col >= 7 and rng.max_row >= 35:
                    range_ok = True
                    print(f"PASS: Component 3 -- Range {rng} covers A2:G35 data area (0.15 pts)")
                    break
            if range_ok:
                break

        if range_ok:
            total_score += 0.15
        else:
            # Partial credit: if range covers at least some data rows
            partial_range = False
            for cf in cf_list:
                has_expression = any(r.type in ('expression', 'formula') for r in cf.rules)
                if not has_expression:
                    continue
                for rng in cf.cells.ranges:
                    if rng.max_row >= 10 and rng.max_col >= 5:
                        partial_range = True
                        break
            if partial_range:
                print(f"PARTIAL: Component 3 -- Range covers some data but not full A2:G35 (0.075 pts)")
                total_score += 0.075
            else:
                print(f"FAIL: Component 3 -- Range does not cover expected data area A2:G35")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: DXF font is bold (0.15 points)
    try:
        bold_found = False
        for cf in cf_list:
            for rule in cf.rules:
                if rule.type in ('expression', 'formula') and rule.dxf and rule.dxf.font:
                    if rule.dxf.font.bold:
                        bold_found = True
                        print(f"PASS: Component 4 -- DXF font is bold (0.15 pts)")
                        break
            if bold_found:
                break
        if not bold_found:
            print("FAIL: Component 4 -- DXF font bold not set in conditional formatting")
        else:
            total_score += 0.15
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # Component 5: DXF font color is dark red (0.15 points)
    try:
        color_ok = False
        found_color = None
        for cf in cf_list:
            for rule in cf.rules:
                if rule.type in ('expression', 'formula') and rule.dxf and rule.dxf.font:
                    font = rule.dxf.font
                    if font.color and font.color.type == 'rgb':
                        rgb_val = font.color.rgb
                        found_color = rgb_val
                        # Dark red variants: 8B0000, 990000, C00000, CC0000, 800000
                        # Accept any reddish dark color: R component high, G and B low
                        if rgb_val:
                            # Strip alpha prefix if 8-char ARGB
                            hex_color = str(rgb_val)
                            if len(hex_color) == 8:
                                hex_color = hex_color[2:]  # strip alpha
                            try:
                                r = int(hex_color[0:2], 16)
                                g = int(hex_color[2:4], 16)
                                b = int(hex_color[4:6], 16)
                                # Dark red: R >= 100, G <= 30, B <= 30
                                if r >= 100 and g <= 30 and b <= 30:
                                    color_ok = True
                                    print(f"PASS: Component 5 -- DXF font color is dark red (RGB: {hex_color}) (0.15 pts)")
                            except (ValueError, IndexError):
                                pass
            if color_ok:
                break

        if color_ok:
            total_score += 0.15
        else:
            print(f"FAIL: Component 5 -- DXF font color not dark red. Found: {found_color}")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
