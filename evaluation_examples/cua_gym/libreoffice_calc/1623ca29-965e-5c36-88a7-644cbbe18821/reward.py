"""
Reward Script: Extract Pending Order IDs using INDEX/SMALL/IF array formulas
Task ID: calc_lf_014
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): E2 contains array formula with INDEX/SMALL/IF
  Component 2 (0.20): E3 contains array formula with INDEX/SMALL/IF
  Component 3 (0.20): E4 contains array formula with INDEX/SMALL/IF
  Component 4 (0.15): All formulas use IFERROR wrapping
  Component 5 (0.15): SMALL k-values are sequential (1, 2, 3)
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_014'


def persist_app_state(domain):
    """Save any unsaved LibreOffice edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def get_formula_text(cell_value):
    """Extract formula text from a cell value, handling ArrayFormula objects."""
    if cell_value is None:
        return None
    # ArrayFormula objects have a .text attribute
    if hasattr(cell_value, 'text'):
        return cell_value.text
    # Regular formula strings start with =
    if isinstance(cell_value, str) and cell_value.startswith('='):
        return cell_value
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Orders' sheet must exist
    if 'Orders' not in wb.sheetnames:
        print("FAIL: 'Orders' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Orders']

    # Extract formula texts from E2, E3, E4
    formulas = {}
    for cell_ref in ['E2', 'E3', 'E4']:
        formulas[cell_ref] = get_formula_text(ws[cell_ref].value)

    # Component 1: E2 contains array formula with INDEX/SMALL/IF (0.30 points)
    try:
        f = formulas.get('E2')
        if f is not None:
            f_upper = f.upper().replace(' ', '')
            has_index = 'INDEX(' in f_upper
            has_small = 'SMALL(' in f_upper
            has_if = 'IF(' in f_upper
            is_array = hasattr(ws['E2'].value, 'text')  # ArrayFormula object

            if has_index and has_small and has_if:
                print(f"PASS: Component 1 — E2 has INDEX/SMALL/IF formula (0.30 pts)")
                print(f"  Formula: {f}")
                total_score += 0.30
            else:
                missing = []
                if not has_index:
                    missing.append('INDEX')
                if not has_small:
                    missing.append('SMALL')
                if not has_if:
                    missing.append('IF')
                print(f"FAIL: Component 1 — E2 missing functions: {missing}")
                print(f"  Formula: {f}")
        else:
            print(f"FAIL: Component 1 — E2 has no formula (value: {repr(ws['E2'].value)})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: E3 contains array formula with INDEX/SMALL/IF (0.20 points)
    try:
        f = formulas.get('E3')
        if f is not None:
            f_upper = f.upper().replace(' ', '')
            has_index = 'INDEX(' in f_upper
            has_small = 'SMALL(' in f_upper
            has_if = 'IF(' in f_upper

            if has_index and has_small and has_if:
                print(f"PASS: Component 2 — E3 has INDEX/SMALL/IF formula (0.20 pts)")
                print(f"  Formula: {f}")
                total_score += 0.20
            else:
                missing = []
                if not has_index:
                    missing.append('INDEX')
                if not has_small:
                    missing.append('SMALL')
                if not has_if:
                    missing.append('IF')
                print(f"FAIL: Component 2 — E3 missing functions: {missing}")
                print(f"  Formula: {f}")
        else:
            print(f"FAIL: Component 2 — E3 has no formula (value: {repr(ws['E3'].value)})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: E4 contains array formula with INDEX/SMALL/IF (0.20 points)
    try:
        f = formulas.get('E4')
        if f is not None:
            f_upper = f.upper().replace(' ', '')
            has_index = 'INDEX(' in f_upper
            has_small = 'SMALL(' in f_upper
            has_if = 'IF(' in f_upper

            if has_index and has_small and has_if:
                print(f"PASS: Component 3 — E4 has INDEX/SMALL/IF formula (0.20 pts)")
                print(f"  Formula: {f}")
                total_score += 0.20
            else:
                missing = []
                if not has_index:
                    missing.append('INDEX')
                if not has_small:
                    missing.append('SMALL')
                if not has_if:
                    missing.append('IF')
                print(f"FAIL: Component 3 — E4 missing functions: {missing}")
                print(f"  Formula: {f}")
        else:
            print(f"FAIL: Component 3 — E4 has no formula (value: {repr(ws['E4'].value)})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: All formulas use IFERROR wrapping (0.15 points)
    try:
        iferror_count = 0
        for cell_ref in ['E2', 'E3', 'E4']:
            f = formulas.get(cell_ref)
            if f is not None and 'IFERROR(' in f.upper():
                iferror_count += 1

        if iferror_count == 3:
            print(f"PASS: Component 4 — All 3 formulas use IFERROR wrapping (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — Only {iferror_count}/3 formulas use IFERROR")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: SMALL k-values are sequential 1, 2, 3 (0.15 points)
    try:
        k_values = []
        for cell_ref in ['E2', 'E3', 'E4']:
            f = formulas.get(cell_ref)
            if f is not None:
                # Extract the k argument from SMALL(..., k)
                # Find all SMALL( calls and extract last argument before closing paren
                f_upper = f.upper().replace(' ', '')
                # Pattern: look for the k value in SMALL(IF(...),k)
                # The k is the last argument before the closing paren of SMALL
                matches = re.findall(r'SMALL\([^)]*\)\s*,\s*(\d+)', f_upper)
                if not matches:
                    # Try alternative pattern where SMALL has nested parens
                    # Find SMALL( then match balanced parens for first arg, then ,k)
                    # Simpler: find all digits right before the closing of the SMALL context
                    match = re.search(r',\s*(\d+)\s*\)\s*\)', f_upper)
                    if match:
                        k_values.append(int(match.group(1)))
                else:
                    k_values.append(int(matches[0]))
            else:
                k_values.append(None)

        if k_values == [1, 2, 3]:
            print(f"PASS: Component 5 — SMALL k-values are sequential [1, 2, 3] (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 5 — k-values are {k_values}, expected [1, 2, 3]")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist app state before verification
persist_app_state("libreoffice_calc")

# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
