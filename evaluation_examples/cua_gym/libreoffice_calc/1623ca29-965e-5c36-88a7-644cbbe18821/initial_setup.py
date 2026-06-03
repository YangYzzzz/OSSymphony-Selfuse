"""
Initial Setup: Extract pending order IDs using array formulas
Task ID: calc_lf_014
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_014'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Orders ---
    ws = wb.active
    ws.title = 'Orders'

    # Headers
    ws['A1'] = 'Order ID'
    ws['B1'] = 'Status'
    ws['C1'] = 'Amount'

    # Data rows
    data = [
        ['ORD-101', 'Shipped',   250],
        ['ORD-102', 'Pending',   180],
        ['ORD-103', 'Delivered', 320],
        ['ORD-104', 'Pending',   145],
        ['ORD-105', 'Pending',   210],
        ['ORD-106', 'Shipped',   175],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Header for pending orders extraction column
    ws['E1'] = 'Pending Orders'

    # E2:E4 left EMPTY -- that is the task target

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
