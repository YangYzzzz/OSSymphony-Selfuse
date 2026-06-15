"""
Reward Script: Fill column E formula down, create column F concatenation summary
Task ID: osworld_calc_formula_pattern_concat_012
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.4 pts): Column E filled down with =C{r}*D{r}/1000 for all 12 data rows
  - Component 2 (0.4 pts): Column F has correct concatenation formulas in F2:F13
  - Component 3 (0.2 pts): Column F header 'Summary' exists in F1
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_formula_pattern_concat_012'


def normalize_formula(f):
    """Normalize a formula string for comparison (strip whitespace, lowercase)."""
    if f is None:
        return ''
    return f.strip().replace(' ', '').lower()


def check_e_formula(formula_str, row):
    """Check if column E formula matches =C{r}*D{r}/1000 pattern."""
    if not isinstance(formula_str, str):
        return False
    # Normalize
    norm = normalize_formula(formula_str)
    # Expected: =c{r}*d{r}/1000
    expected = f'=c{row}*d{row}/1000'
    return norm == expected


def check_f_formula(formula_str, row):
    """
    Check if column F formula matches the required concatenation pattern.
    Expected pattern (case-insensitive, flexible on spacing):
    ="ID: "&A{r}&" | Compound: "&B{r}&" | Temp: "&TEXT(C{r},"0.00")&"degC"&" | Conc: "&TEXT(D{r},"0.00")&" | Rate: "&TEXT(E{r},"0.00")

    The context specifies:
    'ID: [A#] | Compound: [B#] | Temp: [TEXT(C#,"0.00")]°C | Conc: [TEXT(D#,"0.00")] | Rate: [TEXT(E#,"0.00")]'

    After normalization (removing spaces), pipe chars become adjacent to words:
    e.g. " | Compound: " -> "|compound:" in normalized form
    """
    if not isinstance(formula_str, str):
        return False
    norm = normalize_formula(formula_str)
    # Must start with = and be a string formula
    if not norm.startswith('='):
        return False
    r = str(row)
    # Required cell references for the correct row
    cell_refs = [
        f'a{r}',
        f'b{r}',
        f'text(c{r}',
        f'text(d{r}',
        f'text(e{r}',
    ]
    for elem in cell_refs:
        if elem not in norm:
            return False
    # Required keyword labels — after space-removal, separators may be adjacent
    # Check for each label with flexible prefix (space removed leaves pipe or quote adjacent)
    keyword_patterns = [
        'id:',       # starts "id:" right after ="
        'compound:',
        'temp:',
        'conc:',
        'rate:',
        '"0.00"',    # TEXT format string
    ]
    for kw in keyword_patterns:
        if kw not in norm:
            return False
    return True


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task: Fill column E formula down for all 12 data rows, then create column F
    with concatenation formulas combining all data columns with header prefixes.
    """
    total_score = 0.0
    data_rows = list(range(2, 14))  # rows 2-13 (12 data rows)

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Precondition: Verify the file has expected columns (A-E at minimum)
    if ws.max_column < 5:
        print(f"CRITICAL: Expected at least 5 columns, found {ws.max_column}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Column E filled down for all 12 data rows (0.4 points)
    # Initial state: E2 has formula, E3:E13 are None.
    # Golden state: E2:E13 all have =C{r}*D{r}/1000 formulas.
    try:
        e_correct = 0
        e_total = len(data_rows)
        for r in data_rows:
            val = ws.cell(row=r, column=5).value
            if check_e_formula(val, r):
                e_correct += 1
            else:
                print(f"FAIL: E{r} formula check — found: {val!r}")

        if e_correct == e_total:
            print(f"PASS: Component 1 — Column E fully filled down, all {e_total} rows have correct formula (0.4 pts)")
            total_score += 0.4
        elif e_correct > 1:
            # Partial: E2 was already correct in initial state, so only rows 3-13 are task-introduced
            # Only count the newly added rows (3-13). If at least some are filled, partial credit.
            # E2 is pre-existing; rows 3-13 are the task requirement.
            new_rows_correct = 0
            for r in range(3, 14):
                val = ws.cell(row=r, column=5).value
                if check_e_formula(val, r):
                    new_rows_correct += 1
            new_rows_total = 11  # rows 3-13
            if new_rows_correct == new_rows_total:
                print(f"PASS: Component 1 — Column E fully filled down for rows 3-13, all {new_rows_total} new rows correct (0.4 pts)")
                total_score += 0.4
            elif new_rows_correct > 0:
                partial = round(0.4 * (new_rows_correct / new_rows_total), 2)
                print(f"PARTIAL: Component 1 — {new_rows_correct}/{new_rows_total} new E rows have correct formula ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 1 — No new E rows filled. E correct={e_correct}/{e_total}")
        else:
            print(f"FAIL: Component 1 — Column E not filled down. Only {e_correct}/{e_total} rows correct.")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 3: Column F header 'Summary' in F1 (0.2 points)
    # This is checked before Component 2 for logical ordering
    try:
        if ws.max_column >= 6:
            f1_val = ws.cell(row=1, column=6).value
            if f1_val is not None and str(f1_val).strip().lower() == 'summary':
                print(f"PASS: Component 3 — F1 header is 'Summary' (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 3 — F1 header expected 'Summary', found: {f1_val!r}")
        else:
            print(f"FAIL: Component 3 — Column F does not exist (max_col={ws.max_column})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 2: Column F has correct concatenation formulas in F2:F13 (0.4 points)
    # The formula should reference each row's columns A-E with TEXT() formatting for numerics
    # Pattern from context: ID: [A] | Compound: [B] | Temp: [TEXT(C,"0.00")]°C | Conc: [TEXT(D,"0.00")] | Rate: [TEXT(E,"0.00")]
    try:
        if ws.max_column >= 6:
            f_correct = 0
            f_total = len(data_rows)
            for r in data_rows:
                val = ws.cell(row=r, column=6).value
                if check_f_formula(val, r):
                    f_correct += 1
                else:
                    print(f"FAIL: F{r} formula check — found: {val!r}")

            if f_correct == f_total:
                print(f"PASS: Component 2 — Column F has all {f_total} correct concatenation formulas (0.4 pts)")
                total_score += 0.4
            elif f_correct > 0:
                partial = round(0.4 * (f_correct / f_total), 2)
                print(f"PARTIAL: Component 2 — {f_correct}/{f_total} F rows have correct formula ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 2 — No F column formulas found or all incorrect.")
        else:
            print(f"FAIL: Component 2 — Column F does not exist (max_col={ws.max_column})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in given env
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
