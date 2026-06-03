"""
Initial Setup: Research lab data spreadsheet with partial formula in column E
Task ID: osworld_calc_formula_pattern_concat_012
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_formula_pattern_concat_012'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Lab Data ---
    ws = wb.active
    ws.title = 'Lab Data'

    # Headers in row 1
    headers = ['Sample ID', 'Compound', 'Temperature °C', 'Concentration mg/L', 'Reaction Rate']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic research lab data (12 rows)
    # Reaction Rate formula in E2 only: =C2*D2/1000
    # E3:E13 are intentionally left empty (agent must fill them down)
    data = [
        ['S-101', 'Ethanol',          25.0,  150.5],
        ['S-102', 'Acetone',          30.5,  200.0],
        ['S-103', 'Methanol',         22.0,   85.3],
        ['S-104', 'Isopropanol',      28.0,  320.7],
        ['S-105', 'Benzene',          35.5,   42.1],
        ['S-106', 'Toluene',          40.0,  175.9],
        ['S-107', 'Acetic Acid',      20.0,  500.0],
        ['S-108', 'Formic Acid',      23.5,  410.2],
        ['S-109', 'Ethyl Acetate',    32.0,   95.0],
        ['S-110', 'Chloroform',       27.5,  130.4],
        ['S-111', 'Dimethyl Sulfoxide', 38.0, 280.6],
        ['S-112', 'Acetonitrile',     19.5,  360.0],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Add formula ONLY in E2 — agent must fill down E3:E13
    ws['E2'] = '=C2*D2/1000'

    # Column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
