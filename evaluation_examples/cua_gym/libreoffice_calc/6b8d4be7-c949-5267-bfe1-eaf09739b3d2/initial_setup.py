"""
Initial Setup: Employee salary spreadsheet with AutoFilter enabled (no filter applied)
Task ID: calc_dop_filter_aboveavg_011
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_dop_filter_aboveavg_011'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Salaries'

    # --- Headers in row 1 ---
    headers = ['Employee ID', 'Name', 'Department', 'Salary', 'Years', 'Location']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFFFF')
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # --- 55 rows of employee data (rows 2-56) ---
    # Salaries range from $38,000 to $128,000; average ~$72,400
    employees = [
        ('EMP-001', 'Rachel Green',      'Marketing',        58000, 3,  'Chicago'),
        ('EMP-002', 'Ross Geller',       'Research',         92000, 8,  'New York'),
        ('EMP-003', 'Monica Geller',     'Operations',       65000, 5,  'Chicago'),
        ('EMP-004', 'Chandler Bing',     'Finance',          83000, 7,  'New York'),
        ('EMP-005', 'Joey Tribbiani',    'Sales',            42000, 2,  'Los Angeles'),
        ('EMP-006', 'Phoebe Buffay',     'HR',               55000, 4,  'New York'),
        ('EMP-007', 'Sarah Chen',        'Engineering',      89000, 6,  'San Francisco'),
        ('EMP-008', 'Marcus Johnson',    'Engineering',     105000, 10, 'Austin'),
        ('EMP-009', 'Diana Torres',      'Marketing',        48000, 2,  'Miami'),
        ('EMP-010', 'Kevin Park',        'Finance',          75000, 5,  'Chicago'),
        ('EMP-011', 'Lisa Nguyen',       'Research',        103000, 9,  'Seattle'),
        ('EMP-012', 'Tom Bradley',       'Sales',            52000, 3,  'Dallas'),
        ('EMP-013', 'Angela Martinez',   'Operations',       67000, 4,  'Phoenix'),
        ('EMP-014', 'James Wilson',      'Engineering',     110000, 12, 'San Francisco'),
        ('EMP-015', 'Emily Rodriguez',   'HR',               61000, 5,  'New York'),
        ('EMP-016', 'Chris Thompson',    'Finance',          79000, 6,  'Chicago'),
        ('EMP-017', 'Nancy Lee',         'Marketing',        70000, 5,  'Los Angeles'),
        ('EMP-018', 'Robert Kim',        'Research',         86000, 7,  'Boston'),
        ('EMP-019', 'Sophia Davis',      'Engineering',      38000, 1,  'Austin'),
        ('EMP-020', 'Michael Brown',     'Sales',            46000, 2,  'Houston'),
        ('EMP-021', 'Jessica Taylor',    'Operations',       72000, 5,  'Denver'),
        ('EMP-022', 'David White',       'Finance',         105000, 11, 'New York'),
        ('EMP-023', 'Olivia Harris',     'HR',               59000, 4,  'Chicago'),
        ('EMP-024', 'Daniel Jackson',    'Engineering',      82000, 6,  'San Francisco'),
        ('EMP-025', 'Mia Clark',         'Marketing',        44000, 2,  'Atlanta'),
        ('EMP-026', 'Andrew Lewis',      'Research',         97000, 8,  'Seattle'),
        ('EMP-027', 'Chloe Walker',      'Sales',            53000, 3,  'Dallas'),
        ('EMP-028', 'Jason Hall',        'Operations',       72000, 5,  'Denver'),
        ('EMP-029', 'Lauren Allen',      'Engineering',     122000, 14, 'San Francisco'),
        ('EMP-030', 'Ryan Young',        'Finance',          63000, 4,  'Boston'),
        ('EMP-031', 'Amanda Hernandez',  'HR',               57000, 4,  'New York'),
        ('EMP-032', 'Eric King',         'Research',         78000, 6,  'Cambridge'),
        ('EMP-033', 'Stephanie Wright',  'Marketing',        50000, 3,  'Chicago'),
        ('EMP-034', 'Brian Scott',       'Engineering',      87000, 7,  'Austin'),
        ('EMP-035', 'Nicole Green',      'Sales',            40000, 1,  'Miami'),
        ('EMP-036', 'Patrick Baker',     'Operations',      103000, 9,  'Seattle'),
        ('EMP-037', 'Samantha Adams',    'Finance',          74000, 5,  'New York'),
        ('EMP-038', 'Derek Nelson',      'Research',         68000, 4,  'Boston'),
        ('EMP-039', 'Brittany Carter',   'HR',               62000, 4,  'Chicago'),
        ('EMP-040', 'Adam Mitchell',     'Engineering',     128000, 15, 'San Francisco'),
        ('EMP-041', 'Heather Perez',     'Marketing',        47000, 2,  'Los Angeles'),
        ('EMP-042', 'Nathan Roberts',    'Sales',            56000, 3,  'Houston'),
        ('EMP-043', 'Vanessa Turner',    'Operations',       72000, 5,  'Dallas'),
        ('EMP-044', 'Gary Phillips',     'Finance',          84000, 7,  'Chicago'),
        ('EMP-045', 'Tiffany Campbell',  'Research',         88000, 7,  'Seattle'),
        ('EMP-046', 'Kenneth Parker',    'HR',               66000, 4,  'New York'),
        ('EMP-047', 'Pamela Evans',      'Engineering',      73000, 5,  'Austin'),
        ('EMP-048', 'Gregory Edwards',   'Marketing',        54000, 3,  'Atlanta'),
        ('EMP-049', 'Sandra Collins',    'Sales',            43000, 2,  'Dallas'),
        ('EMP-050', 'Raymond Stewart',   'Operations',      108000, 10, 'Chicago'),
        ('EMP-051', 'Carolyn Sanchez',   'Finance',          71000, 5,  'Miami'),
        ('EMP-052', 'Frank Morris',      'Research',         80000, 6,  'Boston'),
        ('EMP-053', 'Maria Rogers',      'HR',               49000, 3,  'Phoenix'),
        ('EMP-054', 'Donald Reed',       'Engineering',      89000, 7,  'San Francisco'),
        ('EMP-055', 'Helen Cook',        'Marketing',        60000, 4,  'Chicago'),
    ]

    for r, row_data in enumerate(employees, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # --- AutoFilter on row 1 (no filter applied) ---
    ws.auto_filter.ref = 'A1:F56'

    # --- Column widths ---
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 16

    # --- Freeze header row ---
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
