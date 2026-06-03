"""
Reward Script: Personal Finance Dashboard
Task ID: calc_wf_043
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25) — KPI formulas in Dashboard row 5 (Total Income, Total Expenses, Net Savings, Savings Rate)
  Component 2 (0.30) — Monthly summary formulas in rows 9-14 (income, expenses with SUMPRODUCT/SUMIFS, net, savings rate)
  Component 3 (0.10) — TOTAL row formulas in row 15
  Component 4 (0.15) — Category breakdown formulas (SUMIF) in rows 19-26
  Component 5 (0.10) — Sparklines/trend indicators in F9:G14
  Component 6 (0.10) — Chart exists on Dashboard sheet
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_043'


def persist_app_state(domain):
    """Try to save any unsaved LibreOffice state."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def is_formula(val):
    """Check if a cell value is a formula string."""
    return isinstance(val, str) and val.startswith('=')


def formula_contains(val, *keywords):
    """Check if formula contains all given keywords (case-insensitive)."""
    if not is_formula(val):
        return False
    upper = val.upper()
    return all(kw.upper() in upper for kw in keywords)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: must have Dashboard sheet
    if 'Dashboard' not in wb.sheetnames:
        print("FAIL: No 'Dashboard' sheet found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Dashboard']

    # Component 1: KPI formulas in row 5 (0.25 points)
    # Golden has: B5=SUMPRODUCT(income), D5=SUMPRODUCT(expenses), F5=B5-D5, H5=savings rate
    # Initial has: all None in row 5
    try:
        kpi_pass = 0
        b5 = ws['B5'].value
        d5 = ws['D5'].value
        f5 = ws['F5'].value
        h5 = ws['H5'].value

        # B5: Total Income formula referencing Transactions amounts
        if is_formula(b5) and formula_contains(b5, 'Transactions'):
            kpi_pass += 1
            print(f"  B5 OK: {b5}")
        else:
            print(f"  B5 FAIL: expected income formula, found: {b5}")

        # D5: Total Expenses formula referencing Transactions amounts
        if is_formula(d5) and formula_contains(d5, 'Transactions'):
            kpi_pass += 1
            print(f"  D5 OK: {d5}")
        else:
            print(f"  D5 FAIL: expected expenses formula, found: {d5}")

        # F5: Net Savings (should reference B5 and D5 or be a formula)
        if is_formula(f5):
            kpi_pass += 1
            print(f"  F5 OK: {f5}")
        else:
            print(f"  F5 FAIL: expected net savings formula, found: {f5}")

        # H5: Savings Rate formula
        if is_formula(h5):
            kpi_pass += 1
            print(f"  H5 OK: {h5}")
        else:
            print(f"  H5 FAIL: expected savings rate formula, found: {h5}")

        comp1_score = 0.25 * (kpi_pass / 4)
        if kpi_pass == 4:
            print(f"PASS: Component 1 — All 4 KPI formulas present ({comp1_score} pts)")
        else:
            print(f"PARTIAL: Component 1 — {kpi_pass}/4 KPI formulas ({comp1_score} pts)")
        total_score += comp1_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Monthly summary formulas in B9:E14 (0.30 points)
    # Golden has: SUMPRODUCT formulas for income (B), expenses (C), net (D), savings rate (E) for 6 months
    # Initial has: all None
    try:
        monthly_formula_count = 0
        total_monthly_cells = 0
        for row_num in range(9, 15):  # rows 9 through 14 (6 months)
            for col_letter in ['B', 'C', 'D', 'E']:
                total_monthly_cells += 1
                cell_val = ws[f'{col_letter}{row_num}'].value
                if is_formula(cell_val):
                    monthly_formula_count += 1

        if total_monthly_cells > 0:
            ratio = monthly_formula_count / total_monthly_cells
            comp2_score = 0.30 * ratio
        else:
            comp2_score = 0.0

        if monthly_formula_count == total_monthly_cells:
            print(f"PASS: Component 2 — All {monthly_formula_count}/{total_monthly_cells} monthly summary formulas present ({comp2_score} pts)")
        elif monthly_formula_count > 0:
            print(f"PARTIAL: Component 2 — {monthly_formula_count}/{total_monthly_cells} monthly formulas ({comp2_score} pts)")
        else:
            print(f"FAIL: Component 2 — No monthly summary formulas found")
        total_score += comp2_score
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: TOTAL row formulas in row 15 (0.10 points)
    # Golden has: A15='TOTAL', B15=SUM(B9:B14), C15=SUM(C9:C14), D15=SUM(D9:D14), E15=savings rate
    # Initial has: A15=None (no totals)
    try:
        total_row_formulas = 0
        for col_letter in ['B', 'C', 'D', 'E']:
            cell_val = ws[f'{col_letter}15'].value
            if is_formula(cell_val):
                total_row_formulas += 1

        # Need at least 3 out of 4 formula cells (B, C, D mandatory, E is savings rate)
        if total_row_formulas >= 3:
            comp3_score = 0.10
            print(f"PASS: Component 3 — TOTAL row has {total_row_formulas}/4 formulas ({comp3_score} pts)")
        elif total_row_formulas > 0:
            comp3_score = 0.10 * (total_row_formulas / 4)
            print(f"PARTIAL: Component 3 — TOTAL row has {total_row_formulas}/4 formulas ({comp3_score} pts)")
        else:
            comp3_score = 0.0
            print(f"FAIL: Component 3 — No TOTAL row formulas in row 15")
        total_score += comp3_score
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Category breakdown formulas in rows 19-26 (0.15 points)
    # Golden has: SUMIF formulas in B19:B26, percentage formulas in C19:C26, avg in D19:D26
    # Initial has: all None in B19:D26
    try:
        cat_formula_count = 0
        total_cat_cells = 0
        for row_num in range(19, 27):  # rows 19 through 26 (8 categories)
            for col_letter in ['B', 'C', 'D']:
                total_cat_cells += 1
                cell_val = ws[f'{col_letter}{row_num}'].value
                if is_formula(cell_val):
                    cat_formula_count += 1

        if total_cat_cells > 0:
            ratio = cat_formula_count / total_cat_cells
            comp4_score = 0.15 * ratio
        else:
            comp4_score = 0.0

        if cat_formula_count == total_cat_cells:
            print(f"PASS: Component 4 — All {cat_formula_count}/{total_cat_cells} category breakdown formulas present ({comp4_score} pts)")
        elif cat_formula_count > 0:
            print(f"PARTIAL: Component 4 — {cat_formula_count}/{total_cat_cells} category formulas ({comp4_score} pts)")
        else:
            print(f"FAIL: Component 4 — No category breakdown formulas found")
        total_score += comp4_score
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Sparklines/trend indicators in F9:G14 (0.10 points)
    # Golden has: REPT formulas in F9:G14 for text-based sparklines
    # Initial has: all None in F9:G14
    try:
        sparkline_count = 0
        total_sparkline_cells = 0
        for row_num in range(9, 15):  # rows 9 through 14
            for col_letter in ['F', 'G']:
                total_sparkline_cells += 1
                cell_val = ws[f'{col_letter}{row_num}'].value
                if is_formula(cell_val):
                    sparkline_count += 1

        if total_sparkline_cells > 0:
            ratio = sparkline_count / total_sparkline_cells
            comp5_score = 0.10 * ratio
        else:
            comp5_score = 0.0

        if sparkline_count == total_sparkline_cells:
            print(f"PASS: Component 5 — All {sparkline_count}/{total_sparkline_cells} sparkline/trend formulas present ({comp5_score} pts)")
        elif sparkline_count > 0:
            print(f"PARTIAL: Component 5 — {sparkline_count}/{total_sparkline_cells} sparkline formulas ({comp5_score} pts)")
        else:
            print(f"FAIL: Component 5 — No sparkline/trend formulas found")
        total_score += comp5_score
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Chart exists on Dashboard (0.10 points)
    # Golden has: 1 BarChart (combo chart for income/expenses/savings rate)
    # Initial has: 0 charts
    try:
        chart_count = len(ws._charts)
        if chart_count >= 1:
            comp6_score = 0.10
            print(f"PASS: Component 6 — {chart_count} chart(s) found on Dashboard ({comp6_score} pts)")
        else:
            comp6_score = 0.0
            print(f"FAIL: Component 6 — No charts found on Dashboard sheet")
        total_score += comp6_score
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
