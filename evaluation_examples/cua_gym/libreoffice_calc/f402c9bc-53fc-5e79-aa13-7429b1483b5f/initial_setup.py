"""
Initial Setup: Personal Finance Dashboard
Task ID: calc_wf_043
Domain: libreoffice_calc

Creates a Transactions sheet with 150 realistic transactions over 6 months,
and a Dashboard sheet with layout headers but NO formulas/charts (pre-task state).
"""

import os
import random
import shlex
import subprocess
import time
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_043'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

random.seed(42)


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def random_date(start, end):
    delta = end - start
    random_days = random.randint(0, delta.days)
    return start + timedelta(days=random_days)


def create_initial():
    wb = openpyxl.Workbook()

    # =========================================================
    # Sheet 1: Transactions
    # =========================================================
    ws_txn = wb.active
    ws_txn.title = 'Transactions'

    headers = ['Date', 'Description', 'Category', 'Amount']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws_txn.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Column widths
    ws_txn.column_dimensions['A'].width = 14
    ws_txn.column_dimensions['B'].width = 35
    ws_txn.column_dimensions['C'].width = 16
    ws_txn.column_dimensions['D'].width = 14

    # Income descriptions by category
    income_items = {
        'Salary': [
            'Monthly salary deposit', 'Bi-weekly paycheck', 'Salary payment',
            'Regular salary transfer', 'Payroll deposit'
        ],
        'Freelance': [
            'Web design project - Meridian Corp', 'Logo redesign for TechStart',
            'Consulting fee - DataFlow Inc', 'UI mockup delivery',
            'SEO audit for GreenLeaf Co', 'Photography session - wedding',
            'Content writing - TravelBlog', 'App prototype - FinanceHub',
            'Social media strategy - Bloom Bakery', 'Copywriting for NovaTech'
        ],
    }

    # Expense descriptions by category
    expense_items = {
        'Rent': [
            'Monthly apartment rent', 'Rent payment - 742 Evergreen Terrace',
            'Housing rent transfer', 'Landlord rent payment'
        ],
        'Food': [
            'Whole Foods Market', 'Trader Joe\'s grocery run', 'DoorDash delivery',
            'Starbucks coffee', 'Chipotle lunch', 'Weekly farmers market',
            'Costco bulk groceries', 'Pizza Hut dinner', 'Thai Palace takeout',
            'Safeway groceries', 'Uber Eats order', 'Panera Bread lunch',
            'Blue Apron subscription', 'Local deli sandwich'
        ],
        'Transport': [
            'Uber ride to downtown', 'Monthly metro pass', 'Gas station fill-up',
            'Lyft to airport', 'Parking garage fee', 'Car insurance payment',
            'Oil change at Jiffy Lube', 'Toll road charges', 'Bus fare'
        ],
        'Entertainment': [
            'Netflix subscription', 'Spotify Premium', 'Movie tickets - AMC',
            'Concert tickets - Madison Square', 'Kindle book purchase',
            'PlayStation Store', 'Bowling night', 'Escape room experience',
            'Museum admission', 'Yoga class drop-in'
        ],
        'Utilities': [
            'Electric bill - ConEdison', 'Water bill', 'Internet - Comcast',
            'Cell phone bill - Verizon', 'Natural gas bill', 'Trash collection fee'
        ],
        'Shopping': [
            'Amazon order - electronics', 'Target household supplies',
            'Nike running shoes', 'IKEA furniture', 'Best Buy headphones',
            'Uniqlo winter jacket', 'Home Depot tools', 'Etsy handmade gifts',
            'Nordstrom sale', 'Apple Store accessories'
        ],
    }

    # Amount ranges
    amount_ranges = {
        'Salary': (4800, 5200),
        'Freelance': (300, 2500),
        'Rent': (1850, 1850),
        'Food': (8, 180),
        'Transport': (5, 120),
        'Entertainment': (10, 85),
        'Utilities': (40, 200),
        'Shopping': (15, 350),
    }

    start_date = datetime(2025, 10, 1)
    end_date = datetime(2026, 3, 31)

    transactions = []

    # Generate salary: 1 per month = 6
    for m in range(6):
        d = datetime(2025, 10, 1) + timedelta(days=30 * m)
        if d.month > 12:
            d = d.replace(year=2026, month=d.month - 12)
        d = d.replace(day=1)
        desc = random.choice(income_items['Salary'])
        amt = round(random.uniform(*amount_ranges['Salary']), 2)
        transactions.append((d, desc, 'Salary', amt))

    # Generate freelance: ~15 total
    for _ in range(15):
        d = random_date(start_date, end_date)
        desc = random.choice(income_items['Freelance'])
        amt = round(random.uniform(*amount_ranges['Freelance']), 2)
        transactions.append((d, desc, 'Freelance', amt))

    # Generate expenses to fill ~129 more rows
    expense_categories = ['Rent', 'Food', 'Transport', 'Entertainment', 'Utilities', 'Shopping']
    expense_weights = [6, 50, 20, 15, 12, 26]  # approximate distribution

    for cat, count in zip(expense_categories, expense_weights):
        for _ in range(count):
            d = random_date(start_date, end_date)
            desc = random.choice(expense_items[cat])
            lo, hi = amount_ranges[cat]
            amt = -round(random.uniform(lo, hi), 2)
            transactions.append((d, desc, cat, amt))

    # Sort by date
    transactions.sort(key=lambda x: x[0])

    # Write transactions
    for r, (dt, desc, cat, amt) in enumerate(transactions, 2):
        ws_txn.cell(row=r, column=1, value=dt).number_format = 'yyyy-mm-dd'
        ws_txn.cell(row=r, column=2, value=desc)
        ws_txn.cell(row=r, column=3, value=cat)
        ws_txn.cell(row=r, column=4, value=amt).number_format = '#,##0.00'

    num_txn = len(transactions)
    print(f'Generated {num_txn} transactions')

    # Freeze header row
    ws_txn.freeze_panes = 'A2'

    # =========================================================
    # Sheet 2: Dashboard (layout only, no formulas/charts)
    # =========================================================
    ws_dash = wb.create_sheet('Dashboard')

    title_font = Font(name='Calibri', size=18, bold=True, color='2F5496')
    subtitle_font = Font(name='Calibri', size=12, bold=True, color='404040')
    kpi_label_font = Font(name='Calibri', size=10, color='808080')
    section_fill = PatternFill(start_color='FFF2F2F2', end_color='FFF2F2F2', fill_type='solid')

    # Title
    ws_dash.merge_cells('A1:H1')
    ws_dash['A1'] = 'Personal Finance Dashboard'
    ws_dash['A1'].font = title_font
    ws_dash['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_dash.row_dimensions[1].height = 36

    # Subtitle
    ws_dash.merge_cells('A2:H2')
    ws_dash['A2'] = 'October 2025 - March 2026'
    ws_dash['A2'].font = Font(name='Calibri', size=11, italic=True, color='808080')
    ws_dash['A2'].alignment = Alignment(horizontal='center')

    # KPI section (row 4-5): labels only, values to be filled by agent
    kpi_labels = ['Total Income', 'Total Expenses', 'Net Savings', 'Savings Rate']
    kpi_cols = [2, 4, 6, 8]  # B, D, F, H
    for label, col in zip(kpi_labels, kpi_cols):
        cell = ws_dash.cell(row=4, column=col, value=label)
        cell.font = kpi_label_font
        cell.alignment = Alignment(horizontal='center')
        # Row 5 is where the KPI values/formulas should go (left empty)
        val_cell = ws_dash.cell(row=5, column=col)
        val_cell.alignment = Alignment(horizontal='center')
        val_cell.font = Font(name='Calibri', size=14, bold=True)

    # Monthly Summary section header (row 7)
    ws_dash.merge_cells('A7:H7')
    ws_dash['A7'] = 'Monthly Summary'
    ws_dash['A7'].font = subtitle_font
    ws_dash['A7'].fill = section_fill
    ws_dash['A7'].alignment = Alignment(horizontal='left', vertical='center')
    ws_dash.row_dimensions[7].height = 24

    # Monthly table headers (row 8)
    monthly_headers = ['Month', 'Income', 'Expenses', 'Net', 'Savings Rate (%)', 'Income Trend', 'Expense Trend']
    month_header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    month_header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    for col, h in enumerate(monthly_headers, 1):
        cell = ws_dash.cell(row=8, column=col, value=h)
        cell.font = month_header_font
        cell.fill = month_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Month labels (rows 9-14) — values/formulas to be added by agent
    months = ['Oct 2025', 'Nov 2025', 'Dec 2025', 'Jan 2026', 'Feb 2026', 'Mar 2026']
    for i, month in enumerate(months):
        ws_dash.cell(row=9 + i, column=1, value=month).font = Font(name='Calibri', size=10)

    # Category Breakdown section header (row 17)
    ws_dash.merge_cells('A17:D17')
    ws_dash['A17'] = 'Category Breakdown'
    ws_dash['A17'].font = subtitle_font
    ws_dash['A17'].fill = section_fill
    ws_dash['A17'].alignment = Alignment(horizontal='left', vertical='center')
    ws_dash.row_dimensions[17].height = 24

    # Category table headers (row 18)
    cat_headers = ['Category', 'Total Spent', '% of Total', 'Avg per Month']
    for col, h in enumerate(cat_headers, 1):
        cell = ws_dash.cell(row=18, column=col, value=h)
        cell.font = month_header_font
        cell.fill = month_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Category labels (rows 19-26) — values to be filled by agent
    categories = ['Salary', 'Freelance', 'Rent', 'Food', 'Transport', 'Entertainment', 'Utilities', 'Shopping']
    for i, cat in enumerate(categories):
        ws_dash.cell(row=19 + i, column=1, value=cat).font = Font(name='Calibri', size=10)

    # Column widths for Dashboard
    ws_dash.column_dimensions['A'].width = 16
    ws_dash.column_dimensions['B'].width = 16
    ws_dash.column_dimensions['C'].width = 16
    ws_dash.column_dimensions['D'].width = 16
    ws_dash.column_dimensions['E'].width = 18
    ws_dash.column_dimensions['F'].width = 16
    ws_dash.column_dimensions['G'].width = 16
    ws_dash.column_dimensions['H'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
