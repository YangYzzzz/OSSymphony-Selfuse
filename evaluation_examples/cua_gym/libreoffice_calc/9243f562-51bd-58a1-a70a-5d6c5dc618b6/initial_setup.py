"""
Initial Setup: E-commerce data for pivot table task
Task ID: calc_pivot_085
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_085'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'EcomData'

    # Headers
    headers = ['OrderID', 'Date', 'Category', 'Product', 'Revenue']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.font = Font(bold=True, size=11, color="FFFFFF")

    # Category config: name -> (target_total, products)
    categories = {
        'Electronics': {
            'total': 85000,
            'products': ['Wireless Headphones', 'Laptop Stand', 'USB-C Hub', 'Bluetooth Speaker',
                         'Smart Watch', 'Tablet Case', 'Portable Charger', 'Webcam HD']
        },
        'Fashion': {
            'total': 62000,
            'products': ['Leather Jacket', 'Running Shoes', 'Silk Scarf', 'Denim Jeans',
                         'Cotton T-Shirt', 'Wool Sweater', 'Canvas Backpack', 'Sunglasses']
        },
        'Home': {
            'total': 48000,
            'products': ['Ceramic Vase', 'Throw Blanket', 'Scented Candle Set', 'Wall Clock',
                         'Kitchen Organizer', 'LED Desk Lamp', 'Doormat', 'Picture Frame Set']
        },
        'Books': {
            'total': 28000,
            'products': ['Python Programming', 'Data Science Handbook', 'Machine Learning Guide',
                         'History of Computing', 'Fiction Bestseller', 'Cooking Essentials',
                         'Travel Photography', 'Business Strategy']
        },
        'Sports': {
            'total': 35000,
            'products': ['Yoga Mat', 'Resistance Bands', 'Jump Rope', 'Water Bottle',
                         'Fitness Tracker', 'Tennis Racket', 'Running Vest', 'Gym Bag']
        },
        'Beauty': {
            'total': 22000,
            'products': ['Face Moisturizer', 'Lip Balm Set', 'Hair Serum', 'Nail Polish Kit',
                         'Eye Cream', 'Body Lotion', 'Makeup Brush Set', 'Perfume Sampler']
        },
    }

    # Distribute 350 rows across categories proportionally to their totals
    grand_total = sum(c['total'] for c in categories.values())  # 280000
    cat_names = list(categories.keys())
    cat_row_counts = {}
    remaining = 350
    for i, name in enumerate(cat_names):
        if i == len(cat_names) - 1:
            cat_row_counts[name] = remaining
        else:
            count = round(350 * categories[name]['total'] / grand_total)
            cat_row_counts[name] = count
            remaining -= count

    # Generate rows for each category with revenues summing to exact target
    all_rows = []
    base_date = datetime(2024, 1, 5)

    for cat_name in cat_names:
        n = cat_row_counts[cat_name]
        target = categories[cat_name]['total']
        products = categories[cat_name]['products']

        # Generate n random values and scale to hit exact target
        raw = [random.uniform(30, 500) for _ in range(n)]
        raw_sum = sum(raw)
        scaled = [round(v * target / raw_sum, 2) for v in raw]
        # Fix rounding error on the last value
        diff = round(target - sum(scaled), 2)
        scaled[-1] = round(scaled[-1] + diff, 2)

        for i in range(n):
            order_date = base_date + timedelta(days=random.randint(0, 364))
            product = random.choice(products)
            all_rows.append((cat_name, order_date, product, scaled[i]))

    # Shuffle to mix categories
    random.shuffle(all_rows)

    # Write data rows
    for idx, (cat, date, product, revenue) in enumerate(all_rows):
        row = idx + 2
        ws.cell(row=row, column=1, value=idx + 1)  # OrderID
        ws.cell(row=row, column=2, value=date).number_format = 'yyyy-mm-dd'
        ws.cell(row=row, column=3, value=cat)
        ws.cell(row=row, column=4, value=product)
        ws.cell(row=row, column=5, value=revenue).number_format = '#,##0.00'

    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = f'A1:E{len(all_rows) + 1}'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Verify totals
    verify_totals(OUTPUT)

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


def verify_totals(path):
    """Quick verification that category totals are correct."""
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb['EcomData']
    totals = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=5):
        cat = row[0].value
        rev = row[2].value
        if cat and rev:
            totals[cat] = totals.get(cat, 0) + rev
    print("Category totals verification:")
    for cat, total in sorted(totals.items()):
        print(f"  {cat}: {total:.2f}")
    print(f"  Grand Total: {sum(totals.values()):.2f}")


create_initial()
