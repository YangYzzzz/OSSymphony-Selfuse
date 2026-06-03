"""
Reward Script: Create pivot table summary and copy static values to StaticReport sheet
Task ID: calc_pivot_085
Domain: libreoffice_calc
Scoring:
  Component 1 (0.15): StaticReport sheet exists with header row
  Component 2 (0.50): Category revenue values correct (6 categories, partial credit each)
  Component 3 (0.15): Grand Total row present and correct
  Component 4 (0.10): Values are static (not formulas)
  Component 5 (0.10): EcomData sheet still intact (350 data rows)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_085'

# Expected category revenue values from task context
EXPECTED_CATEGORIES = {
    'Electronics': 85000,
    'Fashion': 62000,
    'Home': 48000,
    'Books': 28000,
    'Sports': 35000,
    'Beauty': 22000,
}
EXPECTED_GRAND_TOTAL = 280000


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: StaticReport sheet exists with proper header (0.15 points)
    # This FAILS on initial (no StaticReport sheet) and PASSES on golden
    try:
        if 'StaticReport' not in wb.sheetnames:
            print("FAIL: Component 1 — 'StaticReport' sheet does not exist")
            # No StaticReport means nothing else can pass either
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score

        ws_report = wb['StaticReport']

        # Check header row exists
        header_a = ws_report.cell(row=1, column=1).value
        header_b = ws_report.cell(row=1, column=2).value
        if header_a is not None and header_b is not None:
            print(f"PASS: Component 1 — StaticReport sheet exists with headers: "
                  f"'{header_a}', '{header_b}' (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — StaticReport exists but headers missing: "
                  f"A1={header_a}, B1={header_b}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Category revenue values (0.50 points — ~0.083 per category)
    # Each category match earns partial credit
    try:
        # Build a lookup from the StaticReport sheet (skip header row)
        actual_data = {}
        for r in range(2, ws_report.max_row + 1):
            cat = ws_report.cell(row=r, column=1).value
            val = ws_report.cell(row=r, column=2).value
            if cat is not None and str(cat).strip().lower() != 'grand total':
                actual_data[str(cat).strip()] = val

        points_per_cat = round(0.50 / len(EXPECTED_CATEGORIES), 4)
        cat_score = 0.0

        for cat_name, expected_val in EXPECTED_CATEGORIES.items():
            # Try case-insensitive match
            matched_key = None
            for actual_key in actual_data:
                if actual_key.lower() == cat_name.lower():
                    matched_key = actual_key
                    break

            if matched_key is not None:
                actual_val = actual_data[matched_key]
                try:
                    if abs(float(actual_val) - expected_val) < 1.0:
                        print(f"  PASS: {cat_name} = {actual_val} (expected {expected_val})")
                        cat_score += points_per_cat
                    else:
                        print(f"  FAIL: {cat_name} = {actual_val} (expected {expected_val})")
                except (TypeError, ValueError):
                    print(f"  FAIL: {cat_name} value not numeric: {actual_val}")
            else:
                print(f"  FAIL: {cat_name} not found in StaticReport")

        cat_score = round(cat_score, 4)
        if cat_score > 0:
            print(f"PASS: Component 2 — Category values ({cat_score} pts)")
            total_score += cat_score
        else:
            print(f"FAIL: Component 2 — No category values matched")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Grand Total row present and correct (0.15 points)
    try:
        gt_row = None
        for r in range(2, ws_report.max_row + 1):
            cat = ws_report.cell(row=r, column=1).value
            if cat is not None and 'grand' in str(cat).strip().lower() and 'total' in str(cat).strip().lower():
                gt_row = r
                break

        if gt_row is not None:
            val = ws_report.cell(row=gt_row, column=2).value
            try:
                if abs(float(val) - EXPECTED_GRAND_TOTAL) < 1.0:
                    print(f"PASS: Component 3 — Grand Total = {val} (expected {EXPECTED_GRAND_TOTAL}) (0.15 pts)")
                    total_score += 0.15
                else:
                    print(f"FAIL: Component 3 — Grand Total = {val} (expected {EXPECTED_GRAND_TOTAL})")
            except (TypeError, ValueError):
                print(f"FAIL: Component 3 — Grand Total value not numeric: {val}")
        else:
            print("FAIL: Component 3 — Grand Total row not found in StaticReport")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Values are static, not formulas (0.10 points)
    # The task says "paste special > values only" so cells should contain numbers, not formulas
    try:
        formula_count = 0
        for r in range(2, ws_report.max_row + 1):
            val = ws_report.cell(row=r, column=2).value
            if isinstance(val, str) and val.startswith('='):
                formula_count += 1
                print(f"  WARNING: Row {r} col B contains formula: {val}")

        if formula_count == 0:
            print(f"PASS: Component 4 — All revenue values are static (not formulas) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 — {formula_count} values are formulas instead of static values")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: EcomData sheet still intact with 350 data rows (0.10 points)
    # Combined with StaticReport existence — this checks that original data wasn't destroyed
    # AND the new sheet was added. Only scores because StaticReport existing is the gate.
    try:
        if 'EcomData' in wb.sheetnames:
            ws_ecom = wb['EcomData']
            # Check header + 350 data rows = 351 rows
            if ws_ecom.max_row >= 351:
                # Verify a known header
                h1 = ws_ecom.cell(row=1, column=1).value
                if h1 is not None and 'order' in str(h1).lower():
                    print(f"PASS: Component 5 — EcomData intact with {ws_ecom.max_row} rows (0.10 pts)")
                    total_score += 0.10
                else:
                    print(f"FAIL: Component 5 — EcomData header unexpected: {h1}")
            else:
                print(f"FAIL: Component 5 — EcomData has only {ws_ecom.max_row} rows (expected >=351)")
        else:
            print("FAIL: Component 5 — EcomData sheet missing")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
