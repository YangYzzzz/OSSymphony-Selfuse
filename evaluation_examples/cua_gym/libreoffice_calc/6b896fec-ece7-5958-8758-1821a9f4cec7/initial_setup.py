"""
Initial Setup: Create Progress_Tracker spreadsheet with 44 tasks and completion percentages.
Task ID: calc_gcv_048
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_048'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Progress_Tracker"

    # Headers
    ws.cell(row=1, column=1, value="Task Name")
    ws.cell(row=1, column=2, value="Completion %")

    # 44 realistic project tasks with varied completion percentages
    tasks = [
        ("Requirements gathering", 100),
        ("Stakeholder interviews", 85),
        ("Market research analysis", 72),
        ("Technical feasibility study", 100),
        ("Project charter drafting", 90),
        ("Budget estimation", 65),
        ("Resource allocation plan", 48),
        ("Risk assessment matrix", 33),
        ("UI wireframe design", 100),
        ("Database schema design", 78),
        ("API endpoint specification", 55),
        ("Frontend prototype build", 42),
        ("Backend service setup", 60),
        ("Authentication module", 100),
        ("User dashboard development", 37),
        ("Payment integration", 25),
        ("Notification system", 18),
        ("Search functionality", 50),
        ("Reporting module", 0),
        ("Data migration script", 15),
        ("Unit test coverage", 68),
        ("Integration testing", 30),
        ("Performance benchmarking", 12),
        ("Security audit preparation", 45),
        ("Code review process", 80),
        ("CI/CD pipeline setup", 100),
        ("Staging environment config", 92),
        ("Load testing framework", 20),
        ("Documentation portal", 55),
        ("User onboarding flow", 38),
        ("Admin panel features", 62),
        ("Mobile responsive layout", 75),
        ("Accessibility compliance", 28),
        ("Localization support", 10),
        ("Analytics dashboard", 47),
        ("Email template system", 100),
        ("Backup and recovery plan", 82),
        ("Monitoring alerts setup", 58),
        ("Customer feedback module", 0),
        ("Third-party API connectors", 35),
        ("Data export functionality", 70),
        ("Audit trail logging", 100),
        ("Release notes compilation", 88),
        ("Post-launch review prep", 5),
    ]

    for r, (task_name, completion) in enumerate(tasks, 2):
        ws.cell(row=r, column=1, value=task_name)
        ws.cell(row=r, column=2, value=completion)

    # Set column widths for readability
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
