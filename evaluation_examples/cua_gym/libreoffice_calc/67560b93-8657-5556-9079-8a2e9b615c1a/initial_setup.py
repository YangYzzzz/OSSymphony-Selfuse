"""
Initial Setup: Create an advanced XLOOKUP-based salary lookup spreadsheet
Task ID: calc_hr_056
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_056'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: CompPackage ---
    ws1 = wb.active
    ws1.title = 'CompPackage'

    headers = ['Emp ID', 'Name', 'Base', 'Bonus %', 'RSU Value', 'Total Comp']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Employee data
    data = [
        ['E001', 'Alice', 120000, 0.15, 50000],
        ['E002', 'Bob', 95000, 0.10, 25000],
        ['E003', 'Carol', 140000, 0.20, 80000],
        ['E004', 'Dan', 85000, 0.08, 15000],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            if c == 3 or c == 5:  # Base, RSU Value
                cell.number_format = '$#,##0'
            elif c == 4:  # Bonus %
                cell.number_format = '0%'

    # F2:F5 = C*(1+D)+E  (Total Comp formulas)
    for r in range(2, 6):
        cell = ws1.cell(row=r, column=6, value=f'=C{r}*(1+D{r})+E{r}')
        cell.number_format = '$#,##0'

    # Set column widths
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 14
    ws1.column_dimensions['C'].width = 14
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 14
    ws1.column_dimensions['F'].width = 16

    # --- Sheet 2: Lookup ---
    ws2 = wb.create_sheet('Lookup')

    lookup_headers = ['Search ID', 'Name', 'Base', 'Bonus', 'RSU', 'Total']
    for col, h in enumerate(lookup_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', size=11, bold=True)
        cell.alignment = Alignment(horizontal='center')

    # A2 = the search ID
    ws2.cell(row=2, column=1, value='E003')

    # B2:F2 intentionally left EMPTY -- task is to add XLOOKUP formulas here

    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 14
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
