"""
Reward Script: XLOOKUP-based salary lookup for employee compensation package
Task ID: calc_hr_056
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.20): B2 contains lookup formula for Name (target: Carol)
  - Component 2 (0.20): C2 contains lookup formula for Base (target: 140000)
  - Component 3 (0.20): D2 contains lookup formula for Bonus % (target: 0.20)
  - Component 4 (0.20): E2 contains lookup formula for RSU Value (target: 80000)
  - Component 5 (0.20): F2 contains lookup formula for Total Comp (target: 248000)
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_056'


def is_lookup_formula(value):
    """Check if cell value is a lookup formula (XLOOKUP, INDEX-MATCH, VLOOKUP, etc.)."""
    if not isinstance(value, str):
        return False
    upper = value.upper().replace(" ", "")
    # Accept XLOOKUP, INDEX(MATCH(...)), VLOOKUP, HLOOKUP
    return any(kw in upper for kw in ['XLOOKUP', 'INDEX(', 'VLOOKUP', 'HLOOKUP'])


def formula_references_column(formula, col_letter):
    """Check if formula references the expected column in CompPackage."""
    if not isinstance(formula, str):
        return False
    upper = formula.upper().replace(" ", "")
    # For XLOOKUP: check it references CompPackage.<col>2:<col>5 or similar range
    # For INDEX-MATCH: check it references the column letter
    # Accept both sheet reference styles: CompPackage. and CompPackage!
    pattern = rf'COMPPACKAGE[.!]{col_letter}\d'
    return bool(re.search(pattern, upper))


def formula_references_lookup_key(formula):
    """Check if formula uses A2 as the lookup value (the Search ID cell)."""
    if not isinstance(formula, str):
        return False
    upper = formula.upper().replace(" ", "")
    # The formula should reference A2 (the search ID) either directly or via Lookup.A2
    return 'A2' in upper


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Lookup sheet must exist
    if 'Lookup' not in wb.sheetnames:
        print("FAIL: 'Lookup' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Lookup']

    # Precondition: A2 must have 'E003' as the search key
    a2_val = ws['A2'].value
    if a2_val is None or str(a2_val).strip() != 'E003':
        print(f"FAIL: A2 should be 'E003', found: {a2_val}")
        print("REWARD: 0.0")
        return 0.0

    # Define the 5 lookup cells and their expected column references
    # Each cell in B2:F2 should contain a lookup formula that references
    # the corresponding column in CompPackage
    checks = [
        ('B2', 'B', 'Name'),
        ('C2', 'C', 'Base'),
        ('D2', 'D', 'Bonus %'),
        ('E2', 'E', 'RSU Value'),
        ('F2', 'F', 'Total Comp'),
    ]

    for cell_ref, target_col, desc in checks:
        # Component: cell contains a lookup formula referencing correct column (0.20 points)
        try:
            cell_val = ws[cell_ref].value
            if cell_val is None:
                print(f"FAIL: {cell_ref} ({desc}) is empty — no formula present")
                continue

            is_formula = is_lookup_formula(cell_val)
            refs_col = formula_references_column(cell_val, target_col)
            refs_key = formula_references_lookup_key(cell_val)

            if is_formula and refs_col and refs_key:
                print(f"PASS: {cell_ref} ({desc}) — valid lookup formula: {cell_val} (0.20 pts)")
                total_score += 0.20
            elif is_formula and refs_key:
                # Has a lookup formula with correct key but wrong column — partial
                print(f"PARTIAL: {cell_ref} ({desc}) — lookup formula present but column reference unclear: {cell_val} (0.10 pts)")
                total_score += 0.10
            elif is_formula:
                # Has a lookup formula but missing key reference
                print(f"PARTIAL: {cell_ref} ({desc}) — lookup formula present but key/column unclear: {cell_val} (0.05 pts)")
                total_score += 0.05
            else:
                # Cell has a value but it's not a lookup formula
                # Could be a hardcoded value or other formula
                print(f"FAIL: {cell_ref} ({desc}) — not a lookup formula: {cell_val}")
        except Exception as e:
            print(f"ERROR: {cell_ref} ({desc}) — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
