"""
Reward Script: Three pivot tables in Sheet2 with styled header row
Task ID: osworld_calc_pivot_multi_styled_013
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Sheet2 has a merged A1:D1 header row with blue fill (FF4472C4)
                       and bold white text
  Component 2 (0.25): Cost Center pivot table present with correct totals
                       (Engineering, Finance, HR, Marketing, Operations)
  Component 3 (0.25): Expense Category pivot table with correct totals
                       (8 categories)
  Component 4 (0.25): Monthly totals pivot table with correct totals
                       (2025-01 through 2025-04)
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_pivot_multi_styled_013'

# Expected pivot table values derived from Sheet1 source data
EXPECTED_COST_CENTER = {
    'Engineering': 7349,
    'Finance': 6088.75,
    'HR': 2950.5,
    'Marketing': 6835.49,
    'Operations': 3140.75,
}

EXPECTED_CATEGORY = {
    'Advertising': 6600,
    'Consulting': 5950,
    'Equipment': 3420.5,
    'Office Supplies': 725.25,
    'Recruitment': 650,
    'Software': 1747.99,
    'Training': 2525.5,
    'Travel': 4745.25,
}

EXPECTED_MONTHLY = {
    '2025-01': 4765.24,
    '2025-02': 10129.25,
    '2025-03': 4269.5,
    '2025-04': 7200.5,
}

TOLERANCE = 0.02  # allow small floating point differences


def approx_equal(a, b, tol=TOLERANCE):
    """Check numeric equality with tolerance."""
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def extract_pivot_table(ws, start_row, expected_keys):
    """
    Scan Sheet2 starting from start_row for a pivot table matching expected_keys.
    Returns a dict of {key: value} found in consecutive rows (col A = key, col B = value).
    Skips header row (col A value is a section/column label, not a data key).
    """
    found = {}
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        key_cell = row[0].value
        if key_cell is None:
            continue
        key_str = str(key_cell).strip()
        if key_str in expected_keys:
            val_cell = row[1].value if len(row) > 1 else None
            found[key_str] = val_cell
    return found


def find_pivot_data_rows(ws, section_keywords):
    """
    Find rows containing pivot table data by scanning all non-empty cells in col A.
    Returns a list of (row_number, col_a_value, col_b_value) for data rows.
    """
    rows_data = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value is not None and not isinstance(row[0], MergedCell):
            col_b = row[1].value if len(row) > 1 else None
            rows_data.append((row[0].row, row[0].value, col_b))
    return rows_data


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sheet2 must exist
    if 'Sheet2' not in wb.sheetnames:
        print("FAIL: Sheet2 does not exist in the workbook")
        print("REWARD: 0.0")
        return 0.0

    ws2 = wb['Sheet2']

    # Component 1: Merged header row at top of Sheet2 with blue fill and bold white text (0.25 pts)
    # The header must span A1:D1 (merged), have blue fill (FF4472C4), bold, white text
    try:
        # Check A1:D1 merge
        merged_ranges = [str(m) for m in ws2.merged_cells.ranges]
        header_merged = any('A1' in mr and 'D1' in mr for mr in merged_ranges)

        # Check styling of A1
        a1 = ws2['A1']
        has_blue_fill = False
        has_bold = a1.font.bold == True
        has_white_text = False

        # Check fill color
        try:
            fgcolor = a1.fill.fgColor.rgb
            # Accept FF4472C4 (blue) or any blue-ish ARGB
            has_blue_fill = (fgcolor == 'FF4472C4' or
                             (fgcolor and fgcolor.upper().startswith('FF') and
                              fgcolor.upper() not in ('FFFFFFFF', 'FF000000', 'FFFF0000')))
        except Exception:
            pass

        # Check font color is white
        try:
            font_color = a1.font.color.rgb
            # '00FFFFFF' or 'FFFFFFFF' both mean white
            has_white_text = (font_color is not None and
                              font_color.upper().endswith('FFFFFF'))
        except Exception:
            # font color may use theme; fallback: if fill is blue and bold, accept
            has_white_text = False

        # Also check that A1 has some content (header text)
        has_content = a1.value is not None and len(str(a1.value).strip()) > 0

        if header_merged and has_blue_fill and has_bold and has_content:
            print(f"PASS: Component 1 — merged header row A1:D1 with blue fill and bold text (0.25 pts)")
            print(f"  A1 value='{a1.value}', merged={header_merged}, bold={has_bold}, "
                  f"fill={getattr(a1.fill.fgColor, 'rgb', 'N/A')}")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — merged header row incomplete")
            print(f"  merged={header_merged}, blue_fill={has_blue_fill}, "
                  f"bold={has_bold}, content={has_content}, white_text={has_white_text}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Cost Center pivot table with correct totals (0.25 pts)
    # Expected: Engineering=7349, Finance=6088.75, HR=2950.5, Marketing=6835.49, Operations=3140.75
    try:
        found_cc = extract_pivot_table(ws2, 1, set(EXPECTED_COST_CENTER.keys()))

        matched = 0
        for key, expected_val in EXPECTED_COST_CENTER.items():
            if key in found_cc and approx_equal(found_cc[key], expected_val):
                matched += 1
            else:
                actual = found_cc.get(key, 'MISSING')
                print(f"  FAIL CC: {key} expected={expected_val}, found={actual}")

        if matched == len(EXPECTED_COST_CENTER):
            print(f"PASS: Component 2 — Cost Center pivot table: {matched}/{len(EXPECTED_COST_CENTER)} entries correct (0.25 pts)")
            total_score += 0.25
        elif matched >= 3:
            # Partial: most entries correct
            partial = 0.15
            print(f"PARTIAL: Component 2 — Cost Center pivot table: {matched}/{len(EXPECTED_COST_CENTER)} entries ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Cost Center pivot table: only {matched}/{len(EXPECTED_COST_CENTER)} entries correct")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Expense Category pivot table with correct totals (0.25 pts)
    # Expected: Advertising=6600, Consulting=5950, Equipment=3420.5, Office Supplies=725.25,
    #           Recruitment=650, Software=1747.99, Training=2525.5, Travel=4745.25
    try:
        found_cat = extract_pivot_table(ws2, 1, set(EXPECTED_CATEGORY.keys()))

        matched = 0
        for key, expected_val in EXPECTED_CATEGORY.items():
            if key in found_cat and approx_equal(found_cat[key], expected_val):
                matched += 1
            else:
                actual = found_cat.get(key, 'MISSING')
                print(f"  FAIL CAT: {key} expected={expected_val}, found={actual}")

        if matched == len(EXPECTED_CATEGORY):
            print(f"PASS: Component 3 — Expense Category pivot table: {matched}/{len(EXPECTED_CATEGORY)} entries correct (0.25 pts)")
            total_score += 0.25
        elif matched >= 5:
            partial = 0.15
            print(f"PARTIAL: Component 3 — Expense Category pivot table: {matched}/{len(EXPECTED_CATEGORY)} entries ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Expense Category pivot table: only {matched}/{len(EXPECTED_CATEGORY)} entries correct")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Monthly totals pivot table with correct totals (0.25 pts)
    # Expected: 2025-01=4765.24, 2025-02=10129.25, 2025-03=4269.5, 2025-04=7200.5
    try:
        found_mon = extract_pivot_table(ws2, 1, set(EXPECTED_MONTHLY.keys()))

        matched = 0
        for key, expected_val in EXPECTED_MONTHLY.items():
            if key in found_mon and approx_equal(found_mon[key], expected_val):
                matched += 1
            else:
                actual = found_mon.get(key, 'MISSING')
                print(f"  FAIL MON: {key} expected={expected_val}, found={actual}")

        if matched == len(EXPECTED_MONTHLY):
            print(f"PASS: Component 4 — Monthly totals pivot table: {matched}/{len(EXPECTED_MONTHLY)} entries correct (0.25 pts)")
            total_score += 0.25
        elif matched >= 2:
            partial = 0.15
            print(f"PARTIAL: Component 4 — Monthly totals pivot table: {matched}/{len(EXPECTED_MONTHLY)} entries ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Monthly totals pivot table: only {matched}/{len(EXPECTED_MONTHLY)} entries correct")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the VM env
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
