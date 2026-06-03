"""
Initial Setup: Expense report spreadsheet with data in Sheet1
Task ID: osworld_calc_pivot_multi_styled_013
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_pivot_multi_styled_013'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: Expense Report Data ---
    ws1 = wb.active
    ws1.title = 'Sheet1'

    # Column headers
    headers = ['Expense ID', 'Date', 'Cost Center', 'Expense Category', 'Amount', 'Approved By']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # Realistic expense data (20 rows)
    data = [
        ['EXP-1001', '2025-01-08', 'Engineering',  'Travel',          1250.00, 'Rachel Kim'],
        ['EXP-1002', '2025-01-12', 'Marketing',    'Software',        349.99,  'David Park'],
        ['EXP-1003', '2025-01-15', 'HR',            'Training',        875.50,  'Sandra Lee'],
        ['EXP-1004', '2025-01-22', 'Engineering',  'Equipment',       2100.00, 'Rachel Kim'],
        ['EXP-1005', '2025-01-28', 'Finance',       'Office Supplies', 189.75,  'Tom Baker'],
        ['EXP-1006', '2025-02-03', 'Marketing',    'Advertising',     4500.00, 'David Park'],
        ['EXP-1007', '2025-02-07', 'Engineering',  'Software',        799.00,  'Rachel Kim'],
        ['EXP-1008', '2025-02-14', 'HR',            'Recruitment',     650.00,  'Sandra Lee'],
        ['EXP-1009', '2025-02-19', 'Operations',   'Travel',          980.25,  'Chris Wong'],
        ['EXP-1010', '2025-02-25', 'Finance',       'Consulting',      3200.00, 'Tom Baker'],
        ['EXP-1011', '2025-03-05', 'Engineering',  'Training',        450.00,  'Rachel Kim'],
        ['EXP-1012', '2025-03-10', 'Marketing',    'Travel',          1675.00, 'David Park'],
        ['EXP-1013', '2025-03-17', 'Operations',   'Equipment',       1320.50, 'Chris Wong'],
        ['EXP-1014', '2025-03-21', 'HR',            'Office Supplies', 225.00,  'Sandra Lee'],
        ['EXP-1015', '2025-03-28', 'Finance',       'Software',        599.00,  'Tom Baker'],
        ['EXP-1016', '2025-04-04', 'Engineering',  'Consulting',      2750.00, 'Rachel Kim'],
        ['EXP-1017', '2025-04-10', 'Marketing',    'Office Supplies', 310.50,  'David Park'],
        ['EXP-1018', '2025-04-16', 'Operations',   'Travel',          840.00,  'Chris Wong'],
        ['EXP-1019', '2025-04-23', 'HR',            'Training',        1200.00, 'Sandra Lee'],
        ['EXP-1020', '2025-04-30', 'Finance',       'Advertising',     2100.00, 'Tom Baker'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 14
    ws1.column_dimensions['D'].width = 18
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 14

    # Style header row (basic bold only — no heavy formatting)
    header_font = Font(bold=True)
    for col in range(1, 7):
        ws1.cell(row=1, column=col).font = header_font

    # --- Sheet2: Empty (pivot tables to be created by agent) ---
    ws2 = wb.create_sheet('Sheet2')
    # No content — the agent's task is to add pivot tables and styled headers here

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched libreoffice --calc with DISPLAY=:0')


create_initial()
