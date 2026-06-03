"""
Reward Script: Change negative number format from parentheses to minus sign in column D
Task ID: calc_tbl_080
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.5): All data cells in column D use minus-sign format (not parentheses)
  - Component 2 (0.3): Number format string matches expected pattern #,##0;-#,##0
  - Component 3 (0.2): Data values in column D are preserved (not corrupted by format change)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_080'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Find the sheet with column D data (should be 'Transactions' or the active sheet)
    ws = None
    for name in wb.sheetnames:
        candidate = wb[name]
        if candidate['D1'].value is not None:
            ws = candidate
            break
    if ws is None:
        ws = wb.active

    # Determine data rows in column D (skip header row 1)
    data_rows = []
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=4)
        if cell.value is not None:
            data_rows.append(row)

    if not data_rows:
        print("FAIL: No data found in column D")
        print("REWARD: 0.0")
        return 0.0

    print(f"INFO: Found {len(data_rows)} data rows in column D")

    # Component 1: No parentheses format in column D data cells (0.5 points)
    # This checks that the format does NOT use the old parentheses pattern.
    # Initial state has '#,##0;(#,##0)' -> this component FAILS on initial.
    # Golden state has '#,##0;-#,##0' -> this component PASSES on golden.
    try:
        paren_count = 0
        non_paren_count = 0
        for row in data_rows:
            cell = ws.cell(row=row, column=4)
            fmt = cell.number_format or 'General'
            if '(' in fmt and ')' in fmt:
                paren_count += 1
            else:
                non_paren_count += 1

        if paren_count == 0 and non_paren_count == len(data_rows):
            print(f"PASS: Component 1 - All {len(data_rows)} cells in column D have no parentheses format (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 - {paren_count}/{len(data_rows)} cells still use parentheses format")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Column D cells use minus-sign negative format (0.3 points)
    # Specifically checks that the format contains a minus sign section for negatives.
    # Expected format: '#,##0;-#,##0' or similar with '-' in the negative section.
    try:
        minus_format_count = 0
        for row in data_rows:
            cell = ws.cell(row=row, column=4)
            fmt = cell.number_format or 'General'
            # The format should have a semicolon-separated negative section with a minus sign
            parts = fmt.split(';')
            if len(parts) >= 2:
                neg_part = parts[1].strip()
                if '-' in neg_part:
                    minus_format_count += 1

        if minus_format_count == len(data_rows):
            print(f"PASS: Component 2 - All {len(data_rows)} cells use minus-sign negative format (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 - Only {minus_format_count}/{len(data_rows)} cells have minus-sign format")
            # Sample a failing cell for debugging
            for row in data_rows:
                cell = ws.cell(row=row, column=4)
                fmt = cell.number_format or 'General'
                parts = fmt.split(';')
                if len(parts) < 2 or '-' not in parts[1]:
                    print(f"  Example: D{row} format = {fmt!r}")
                    break
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Data values preserved correctly (0.2 points)
    # This is a compound check: data integrity AND format change happened.
    # We verify that numeric values are intact AND that the format is not 'General'
    # (which would mean the format was removed rather than changed).
    # The 'not General' sub-condition ensures this fails on initial only if
    # combined with the format-change requirement.
    try:
        expected_values = {
            2: 12500, 3: -3200, 4: -875, 5: 8400, 6: -15600,
            7: -2450, 8: 6320, 9: -4100, 10: -1890, 11: 9750,
            12: -520, 13: -1340, 14: -2800, 15: -500, 16: 4200
        }
        values_ok = 0
        total_checked = 0
        format_changed_count = 0
        for row, expected in expected_values.items():
            cell = ws.cell(row=row, column=4)
            if cell.value is not None:
                total_checked += 1
                try:
                    if abs(float(cell.value) - expected) < 0.01:
                        values_ok += 1
                except (ValueError, TypeError):
                    pass
            # Check that format is not parentheses (i.e., change happened)
            fmt = cell.number_format or 'General'
            if ';' in fmt and '(' not in fmt and '-' in fmt:
                format_changed_count += 1

        if total_checked > 0 and values_ok == total_checked and format_changed_count > 0:
            print(f"PASS: Component 3 - All {values_ok} values preserved and format changed (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 - Values OK: {values_ok}/{total_checked}, format_changed_count: {format_changed_count}")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    final_score = round(min(total_score, 1.0), 1)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
