"""
Initial Setup: Spreadsheet with custom number format using parentheses for negatives
Task ID: calc_tbl_080
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_080'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Transactions ---
    ws = wb.active
    ws.title = "Transactions"

    # Headers
    headers = ["Date", "Description", "Category", "Balance"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Transaction data with mix of positive and negative balances
    data = [
        ["2025-01-05", "Client payment - Apex Corp", "Revenue", 12500],
        ["2025-01-08", "Office rent - January", "Facilities", -3200],
        ["2025-01-10", "Software licenses renewal", "IT", -875],
        ["2025-01-12", "Consulting fee - Rivera & Associates", "Revenue", 8400],
        ["2025-01-15", "Employee payroll - Bi-weekly", "Payroll", -15600],
        ["2025-01-17", "Marketing campaign - Q1 launch", "Marketing", -2450],
        ["2025-01-20", "Product sales - Online store", "Revenue", 6320],
        ["2025-01-22", "Insurance premium - Annual", "Insurance", -4100],
        ["2025-01-25", "Equipment purchase - Monitors", "IT", -1890],
        ["2025-01-27", "Client payment - Meridian LLC", "Revenue", 9750],
        ["2025-01-28", "Utilities - Electric & Water", "Facilities", -520],
        ["2025-01-29", "Travel expenses - Conference", "Travel", -1340],
        ["2025-01-30", "Freelancer payment - Design work", "Contractor", -2800],
        ["2025-01-31", "Refund - Returned merchandise", "Revenue", -500],
        ["2025-02-01", "Subscription revenue - Monthly", "Revenue", 4200],
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        ws.cell(row=r, column=2, value=row_data[1])
        ws.cell(row=r, column=3, value=row_data[2])
        ws.cell(row=r, column=4, value=row_data[3])

    # Apply parentheses format to Column D (the key initial state)
    paren_format = '#,##0;(#,##0)'
    for r in range(2, len(data) + 2):
        ws.cell(row=r, column=4).number_format = paren_format

    # Set column widths for readability
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14

    # Freeze header row
    ws.freeze_panes = "A2"

    # --- Sheet 2: Summary ---
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Category"
    ws2["B1"] = "Total"
    ws2["A1"].font = Font(bold=True)
    ws2["B1"].font = Font(bold=True)

    categories = ["Revenue", "Facilities", "IT", "Payroll", "Marketing",
                   "Insurance", "Travel", "Contractor"]
    for i, cat in enumerate(categories, 2):
        ws2.cell(row=i, column=1, value=cat)
        ws2.cell(row=i, column=2, value=f'=SUMIF(Transactions!C:C,A{i},Transactions!D:D)')
        ws2.cell(row=i, column=2).number_format = paren_format

    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
