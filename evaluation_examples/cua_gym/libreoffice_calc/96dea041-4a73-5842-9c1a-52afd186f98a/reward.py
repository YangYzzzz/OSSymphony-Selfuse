"""
Reward Script: Procurement Vendor Evaluation Scorecard
Task ID: calc_grs_072
Domain: libreoffice_calc
Scoring:
  C1 (0.25) - SUMPRODUCT weighted score formulas in J6:J11
  C2 (0.15) - RANK formulas in K6:K11
  C3 (0.15) - Conditional formatting with gold fill for top-ranked vendor
  C4 (0.20) - Radar chart comparing top 3 vendors
  C5 (0.15) - Comments section filled with qualitative notes
  C6 (0.10) - Tiebreaker rules documented in A18:A19
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_072'


def persist_app_state(domain: str):
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Scorecard' not in wb.sheetnames:
        print("CRITICAL: 'Scorecard' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Scorecard']

    # Component 1: SUMPRODUCT weighted score formulas in J6:J11 (0.25 points)
    # The golden has =SUMPRODUCT(B6:I6,$B$20:$I$20) style formulas
    # Initial has J6:J11 all empty — this is a task-introduced change
    try:
        sumproduct_count = 0
        for r in range(6, 12):
            cell_val = ws.cell(row=r, column=10).value  # column J
            if cell_val is not None and isinstance(cell_val, str):
                # Accept SUMPRODUCT formulas in any reasonable form
                normalized = cell_val.upper().replace(" ", "")
                if "SUMPRODUCT" in normalized:
                    sumproduct_count += 1
        if sumproduct_count == 6:
            print(f"PASS: Component 1 — All 6 SUMPRODUCT formulas found in J6:J11 (0.25 pts)")
            total_score += 0.25
        elif sumproduct_count >= 3:
            partial = 0.25 * (sumproduct_count / 6)
            print(f"PARTIAL: Component 1 — {sumproduct_count}/6 SUMPRODUCT formulas found ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Expected SUMPRODUCT formulas in J6:J11, found {sumproduct_count}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: RANK formulas in K6:K11 (0.15 points)
    # Initial has K6:K11 all empty
    try:
        rank_count = 0
        for r in range(6, 12):
            cell_val = ws.cell(row=r, column=11).value  # column K
            if cell_val is not None and isinstance(cell_val, str):
                normalized = cell_val.upper().replace(" ", "")
                if "RANK" in normalized:
                    rank_count += 1
        if rank_count == 6:
            print(f"PASS: Component 2 — All 6 RANK formulas found in K6:K11 (0.15 pts)")
            total_score += 0.15
        elif rank_count >= 3:
            partial = 0.15 * (rank_count / 6)
            print(f"PARTIAL: Component 2 — {rank_count}/6 RANK formulas found ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Expected RANK formulas in K6:K11, found {rank_count}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Conditional formatting with gold fill for top-ranked vendor (0.15 points)
    # Initial has no conditional formatting rules
    try:
        cf_rules = list(ws.conditional_formatting)
        gold_cf_found = False
        for cf in cf_rules:
            for rule in cf.rules:
                # Check for an expression-based or formula-based rule
                # The golden uses expression type with formula '$K6=1' and gold fill FFFFD700
                if rule.dxf and rule.dxf.fill:
                    try:
                        fill_rgb = rule.dxf.fill.fgColor.rgb if rule.dxf.fill.fgColor else None
                        if fill_rgb:
                            # Check for gold-ish colors (FFD700 is standard gold)
                            rgb_lower = fill_rgb.lower()
                            if 'ffd700' in rgb_lower or 'ffc000' in rgb_lower or 'ffbf00' in rgb_lower:
                                gold_cf_found = True
                    except Exception:
                        pass
        if gold_cf_found:
            print(f"PASS: Component 3 — Conditional formatting with gold fill found (0.15 pts)")
            total_score += 0.15
        elif len(cf_rules) > 0:
            # Some conditional formatting exists but not gold specifically
            print(f"PARTIAL: Component 3 — Conditional formatting found but no gold fill detected (0.07 pts)")
            total_score += 0.07
        else:
            print(f"FAIL: Component 3 — No conditional formatting rules found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Radar chart comparing top 3 vendors (0.20 points)
    # Initial has 0 charts
    try:
        chart_count = len(ws._charts)
        radar_found = False
        correct_series_count = False
        if chart_count > 0:
            for chart in ws._charts:
                chart_class = chart.__class__.__name__
                if 'Radar' in chart_class:
                    radar_found = True
                    if len(chart.series) >= 3:
                        correct_series_count = True

        if radar_found and correct_series_count:
            print(f"PASS: Component 4 — Radar chart with 3+ vendor series found (0.20 pts)")
            total_score += 0.20
        elif radar_found:
            print(f"PARTIAL: Component 4 — Radar chart found but fewer than 3 series (0.12 pts)")
            total_score += 0.12
        elif chart_count > 0:
            # Some chart exists but it's not a radar chart
            print(f"PARTIAL: Component 4 — Chart found but not radar type (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 4 — No charts found in Scorecard sheet")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Comments section filled with qualitative notes (0.15 points)
    # Initial Comments sheet has vendor names in col A but cols B-I are all empty
    try:
        if 'Comments' not in wb.sheetnames:
            print(f"FAIL: Component 5 — 'Comments' sheet not found")
        else:
            ws_comments = wb['Comments']
            filled_cells = 0
            total_cells = 0
            for r in range(4, 10):  # 6 vendors in rows 4-9
                for c in range(2, 10):  # columns B through I (8 criteria)
                    total_cells += 1
                    val = ws_comments.cell(row=r, column=c).value
                    if val is not None and str(val).strip() != '':
                        filled_cells += 1

            fill_ratio = filled_cells / total_cells if total_cells > 0 else 0
            if fill_ratio >= 0.5:
                print(f"PASS: Component 5 — Comments filled: {filled_cells}/{total_cells} cells ({fill_ratio:.0%}) (0.15 pts)")
                total_score += 0.15
            elif fill_ratio > 0:
                partial = 0.15 * fill_ratio
                print(f"PARTIAL: Component 5 — Comments partially filled: {filled_cells}/{total_cells} cells ({partial:.2f} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 5 — Comments sheet has no qualitative notes (0/{total_cells} cells filled)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Tiebreaker rules documented with actual content (0.10 points)
    # Initial has "Tiebreaker Rules:" label in A17 but A18 and A19 are empty
    # Golden has actual rule text describing specific tiebreaker criteria in A18:A19
    # We check for substantive rule content (not just the label)
    try:
        tiebreaker_content_found = False
        # Search rows 17-25 for substantive tiebreaker rule content
        # Must contain actual criteria references (quality, delivery, price, etc.)
        for r in range(17, 26):
            for c in range(1, 12):
                val = ws.cell(row=r, column=c).value
                if val and isinstance(val, str) and len(val) > 30:
                    val_lower = val.lower()
                    # Must reference specific tiebreaker criteria
                    criteria_keywords = ['quality', 'delivery', 'price', 'competitiveness']
                    tie_keywords = ['tied', 'tie', 'tiebreaker', 'same weighted', 'share the same', 'precedence']
                    has_criteria = any(kw in val_lower for kw in criteria_keywords)
                    has_tie_ref = any(kw in val_lower for kw in tie_keywords)
                    if has_criteria and has_tie_ref:
                        tiebreaker_content_found = True
                        break
            if tiebreaker_content_found:
                break

        if tiebreaker_content_found:
            print(f"PASS: Component 6 — Tiebreaker rules with specific criteria documented (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 6 — No substantive tiebreaker rules found (need specific criteria references)")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist any unsaved GUI state before verification
persist_app_state("libreoffice_calc")

# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
