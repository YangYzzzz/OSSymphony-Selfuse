"""
Initial Setup: Procurement Vendor Evaluation Scorecard
Task ID: calc_grs_072
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_072'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ── Sheet 1: Scorecard ──
    ws = wb.active
    ws.title = "Scorecard"

    # Title
    ws.merge_cells("A1:K1")
    ws["A1"] = "Procurement Vendor Evaluation Scorecard — 2-Year Contract ($500,000)"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Subtitle
    ws.merge_cells("A2:K2")
    ws["A2"] = "Evaluation Period: Q2 2026 | Contract Award Target: April 21, 2026"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="4472C4")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Row 4: Headers
    headers = [
        "Vendor",
        "Price\nCompetitiveness",
        "Quality",
        "Delivery\nReliability",
        "Customer\nService",
        "Financial\nStability",
        "Technical\nCapability",
        "Sustainability\n/ ESG",
        "References",
        "Weighted\nScore",
        "Rank",
    ]
    header_fill = PatternFill(start_color="FF1F3864", end_color="FF1F3864", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    # Row 5: Weights
    weights = [
        "Weight →",
        "25%", "20%", "20%", "15%", "10%", "5%", "3%", "2%",
        "", "",
    ]
    weight_fill = PatternFill(start_color="FFD6DCE4", end_color="FFD6DCE4", fill_type="solid")
    weight_font = Font(name="Arial", size=10, bold=True)
    for col, w in enumerate(weights, 1):
        cell = ws.cell(row=5, column=col, value=w)
        cell.fill = weight_fill
        cell.font = weight_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Row 5 weights as numbers in a hidden helper row? No — keep it simple.
    # Actual numeric weights will go in a helper area (row 3 hidden or row 20+).

    # Numeric weights in row 20 (for SUMPRODUCT reference) — label them
    ws.cell(row=20, column=1, value="Numeric Weights (for formulas)")
    ws.cell(row=20, column=1).font = Font(italic=True, color="808080", size=9)
    numeric_weights = [0.25, 0.20, 0.20, 0.15, 0.10, 0.05, 0.03, 0.02]
    for col, nw in enumerate(numeric_weights, 2):
        cell = ws.cell(row=20, column=col, value=nw)
        cell.number_format = '0%'
        cell.font = Font(color="808080", size=9)

    # Vendors and scores (rows 6-11)
    vendors = [
        {
            "name": "Apex Industrial Solutions",
            "scores": [4, 5, 4, 3, 5, 4, 3, 4],
        },
        {
            "name": "BlueWave Supply Co.",
            "scores": [5, 3, 3, 4, 4, 3, 4, 3],
        },
        {
            "name": "Cornerstone Materials Ltd.",
            "scores": [3, 4, 5, 5, 3, 5, 4, 5],
        },
        {
            "name": "Dynamo Procurement Group",
            "scores": [4, 4, 4, 4, 4, 3, 2, 3],
        },
        {
            "name": "EverGreen Sourcing Inc.",
            "scores": [2, 5, 5, 4, 3, 4, 5, 4],
        },
        {
            "name": "FairTrade Logistics Partners",
            "scores": [3, 3, 3, 5, 5, 4, 5, 5],
        },
    ]

    data_font = Font(name="Arial", size=10)
    score_align = Alignment(horizontal="center", vertical="center")
    vendor_align = Alignment(horizontal="left", vertical="center")

    # Alternating row colors for readability
    light_fill = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")

    for i, v in enumerate(vendors):
        row = 6 + i
        row_fill = light_fill if i % 2 == 0 else white_fill

        # Vendor name
        cell = ws.cell(row=row, column=1, value=v["name"])
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.alignment = vendor_align
        cell.border = thin_border
        cell.fill = row_fill

        # Scores (columns 2-9)
        for col, score in enumerate(v["scores"], 2):
            cell = ws.cell(row=row, column=col, value=score)
            cell.font = data_font
            cell.alignment = score_align
            cell.border = thin_border
            cell.fill = row_fill

        # Weighted Score column (J / col 10) — EMPTY in initial
        cell = ws.cell(row=row, column=10)
        cell.border = thin_border
        cell.fill = row_fill

        # Rank column (K / col 11) — EMPTY in initial
        cell = ws.cell(row=row, column=11)
        cell.border = thin_border
        cell.fill = row_fill

    # Column widths
    ws.column_dimensions["A"].width = 32
    for col_letter in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        ws.column_dimensions[col_letter].width = 14
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 10

    # Row heights
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[4].height = 36
    ws.row_dimensions[5].height = 20

    # Scoring legend
    ws.cell(row=14, column=1, value="Scoring Scale:")
    ws.cell(row=14, column=1).font = Font(bold=True, size=10)
    ws.cell(row=15, column=1, value="1 = Poor  |  2 = Below Average  |  3 = Average  |  4 = Good  |  5 = Excellent")
    ws.cell(row=15, column=1).font = Font(size=9, italic=True)

    # Tiebreaker section header — empty, to be filled in golden
    ws.cell(row=17, column=1, value="Tiebreaker Rules:")
    ws.cell(row=17, column=1).font = Font(bold=True, size=10)
    # Leave rows 18-19 empty for tiebreaker content

    # Freeze header row
    ws.freeze_panes = "A6"

    # ── Sheet 2: Comments ──
    ws2 = wb.create_sheet("Comments")
    ws2.merge_cells("A1:I1")
    ws2["A1"] = "Qualitative Evaluation Notes — Vendor Assessment"
    ws2["A1"].font = Font(name="Arial", size=13, bold=True, color="1F3864")
    ws2["A1"].alignment = Alignment(horizontal="center")

    # Headers: Vendor | then 8 criteria
    criteria_names = [
        "Price Competitiveness", "Quality", "Delivery Reliability",
        "Customer Service", "Financial Stability", "Technical Capability",
        "Sustainability/ESG", "References",
    ]
    comment_headers = ["Vendor"] + criteria_names
    for col, h in enumerate(comment_headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Vendor names in column A (rows 4-9)
    for i, v in enumerate(vendors):
        cell = ws2.cell(row=4 + i, column=1, value=v["name"])
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.border = thin_border
        # Leave criteria columns empty for comments
        for col in range(2, 10):
            ws2.cell(row=4 + i, column=col).border = thin_border

    ws2.column_dimensions["A"].width = 32
    for col_letter in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        ws2.column_dimensions[col_letter].width = 22
    ws2.row_dimensions[3].height = 36

    # Save
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
