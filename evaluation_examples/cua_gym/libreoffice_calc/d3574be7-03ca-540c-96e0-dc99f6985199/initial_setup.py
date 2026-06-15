"""
Initial Setup: Podcast subscription tracker spreadsheet
Task ID: osworld_multi_apps_misc_019
Domain: libreoffice_calc (multi-app: Chrome + LibreOffice Calc)

Creates my_podcasts.xlsx with a sheet of podcasts the user is already subscribed to.
Chrome should be opened to the Spotify podcast charts page.
The user does NOT yet have a 'new_podcasts' sheet.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_misc_019'
OUTPUT = f'{WORKDIR}/my_podcasts.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Subscribed Podcasts ---
    ws = wb.active
    ws.title = 'my_podcasts'

    # Headers
    headers = ['Rank', 'Title', 'Host', 'Category']
    header_font = Font(bold=True, name='Calibri', size=11)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11, color='FFFFFFFF')
        cell.fill = header_fill
        cell.alignment = header_align

    # User's existing subscribed podcasts (some from top charts, various years)
    # These are podcasts the user already follows - they should NOT appear in new_podcasts
    # Mix of pre-2023 and post-2023 shows, some popular on Spotify charts
    data = [
        [1,  'The Daily',                         'Michael Barbaro',     'News'],
        [2,  'Crime Junkie',                       'Ashley Flowers',      'True Crime'],
        [3,  'Serial',                             'Sarah Koenig',        'True Crime'],
        [4,  'Stuff You Should Know',              'Josh Clark & Chuck Bryant', 'Education'],
        [5,  'Conan O\'Brien Needs a Friend',      'Conan O\'Brien',      'Comedy'],
        [6,  'Radiolab',                           'Lulu Miller & Latif Nasser', 'Science'],
        [7,  'Fresh Air',                          'Terry Gross',         'Culture'],
        [8,  'TED Talks Daily',                    'Various Hosts',       'Education'],
        [9,  'My Favorite Murder',                 'Karen Kilgariff & Georgia Hardstark', 'True Crime'],
        [10, 'Hidden Brain',                       'Shankar Vedantam',    'Psychology'],
        [11, 'SmartLess',                          'Jason Bateman, Sean Hayes & Will Arnett', 'Comedy'],
        [12, 'Pod Save America',                   'Jon Favreau & Friends', 'Politics'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name='Calibri', size=11)
            if r % 2 == 0:
                cell.fill = PatternFill(start_color='FFE9EFF7', end_color='FFE9EFF7', fill_type='solid')

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 16

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open Chrome to the Spotify podcast charts page first
    launch_gui(
        'google-chrome --new-window "https://chartable.com/charts/spotify"',
        delay_sec=3.0
    )

    # Then open LibreOffice Calc with the podcasts file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)

    print('GUI_READY: launched Chrome (chartable.com/charts/spotify) and LibreOffice Calc with DISPLAY=:0')


create_initial()
