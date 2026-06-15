"""
Reward Script: Check Spotify podcast charts and add new sheet with unsubscribed pre-2023 podcasts
Task ID: osworld_multi_apps_misc_019
Domain: libreoffice_calc
Scoring:
  - Component 1: 'new_podcasts' sheet exists                          (0.30 pts)
  - Component 2: Correct headers (Rank, Title, Host, Category)       (0.20 pts)
  - Component 3: Entries are sorted by Rank ascending                (0.20 pts)
  - Component 4: No entry overlaps with 'my_podcasts' subscriptions  (0.15 pts)
  - Component 5: All rank values are in range 1-20 (top 20 check)    (0.15 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_misc_019'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: 'new_podcasts' sheet exists (0.30 points)
    # This is the primary task change — a new sheet was created
    try:
        if 'new_podcasts' in wb.sheetnames:
            print("PASS: Component 1 — 'new_podcasts' sheet exists (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 — 'new_podcasts' sheet NOT found. Sheets: {wb.sheetnames}")
            # Cannot proceed further without the sheet
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Access the new_podcasts sheet
    ws_new = wb['new_podcasts']

    # Component 2: Correct headers (Rank, Title, Host, Category) (0.20 points)
    # Headers must match the existing 'my_podcasts' sheet exactly
    try:
        expected_headers = ['Rank', 'Title', 'Host', 'Category']
        actual_headers = []
        for col in range(1, 5):
            val = ws_new.cell(row=1, column=col).value
            actual_headers.append(val)

        # Check all four expected headers are present
        headers_match = all(
            actual_headers[i] == expected_headers[i]
            for i in range(len(expected_headers))
            if i < len(actual_headers)
        )
        if headers_match and len(actual_headers) >= 4:
            print(f"PASS: Component 2 — headers correct: {actual_headers} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 — expected headers {expected_headers}, found {actual_headers}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Gather all data rows from new_podcasts (rows 2 onwards)
    new_podcast_rows = []
    try:
        for row in ws_new.iter_rows(min_row=2, max_row=ws_new.max_row, values_only=True):
            # Only include rows that have at least a rank value
            if row[0] is not None:
                new_podcast_rows.append(row)
    except Exception as e:
        print(f"ERROR: Gathering rows — {e}")

    # Component 3: Entries are sorted by Rank ascending (0.20 points)
    # Ranks in new_podcasts should be in ascending order
    try:
        if len(new_podcast_rows) > 0:
            ranks = [r[0] for r in new_podcast_rows if r[0] is not None]
            is_sorted = all(ranks[i] <= ranks[i+1] for i in range(len(ranks)-1))
            if is_sorted:
                print(f"PASS: Component 3 — entries sorted by Rank ascending: {ranks} (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 3 — ranks not in ascending order: {ranks}")
        else:
            print("FAIL: Component 3 — no data rows found in 'new_podcasts'")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: No entries overlap with 'my_podcasts' subscriptions (0.15 points)
    # The task says: find podcasts I'm NOT already subscribed to
    try:
        if 'my_podcasts' in wb.sheetnames:
            ws_existing = wb['my_podcasts']
            # Collect titles from existing subscriptions (case-insensitive)
            existing_titles = set()
            for row in ws_existing.iter_rows(min_row=2, max_row=ws_existing.max_row, values_only=True):
                if row[1] is not None:
                    existing_titles.add(str(row[1]).strip().lower())

            # Check that none of the new_podcasts titles are already in my_podcasts
            new_titles = [str(r[1]).strip().lower() for r in new_podcast_rows if r[1] is not None]
            overlap = [t for t in new_titles if t in existing_titles]

            if len(overlap) == 0 and len(new_podcast_rows) > 0:
                print(f"PASS: Component 4 — no overlap with existing subscriptions (0.15 pts)")
                total_score += 0.15
            elif len(overlap) > 0:
                print(f"FAIL: Component 4 — {len(overlap)} entries overlap with my_podcasts: {overlap}")
            else:
                print("FAIL: Component 4 — no data rows to check for overlap")
        else:
            print("FAIL: Component 4 — 'my_podcasts' sheet not found (cannot check overlap)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: All rank values are from the Top 20 (1-20 range) (0.15 points)
    # The task specifies "Top 20 podcasts" from Spotify charts
    try:
        if len(new_podcast_rows) > 0:
            ranks = [r[0] for r in new_podcast_rows if r[0] is not None]
            all_in_top20 = all(isinstance(r, int) and 1 <= r <= 20 for r in ranks)
            has_data = len(ranks) > 0

            if all_in_top20 and has_data:
                print(f"PASS: Component 5 — all {len(ranks)} entries have ranks in 1-20 range (0.15 pts)")
                total_score += 0.15
            else:
                out_of_range = [r for r in ranks if not (isinstance(r, int) and 1 <= r <= 20)]
                print(f"FAIL: Component 5 — ranks out of 1-20 range: {out_of_range}")
        else:
            print("FAIL: Component 5 — no data rows to check rank range")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the VM environment
file_path = f'{WORKDIR}/my_podcasts.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
