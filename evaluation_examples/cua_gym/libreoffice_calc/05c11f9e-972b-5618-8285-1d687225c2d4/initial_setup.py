"""
Initial Setup: Competitive Analysis Matrix
Task ID: calc_wf_061
Domain: libreoffice_calc

Creates a spreadsheet with 15 evaluation criteria across 4 categories,
7 companies (own + 6 competitors), raw scores 1-10, category weights.
Summary section at bottom with labels but NO formulas, NO charts,
NO conditional formatting, NO SWOT text.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_061'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analysis"

    # --- Company names (columns C-I) ---
    companies = [
        "NovaTech Solutions",  # Our company
        "Apex Industries",
        "Pinnacle Corp",
        "Meridian Systems",
        "Vanguard Digital",
        "Summit Enterprises",
        "Catalyst Group",
    ]

    # --- Category definitions ---
    # (category_name, [(criteria_name, weight), ...])
    categories = [
        ("Product", [
            ("Feature Completeness", 0.25),
            ("Ease of Use", 0.20),
            ("Reliability", 0.20),
            ("Innovation", 0.15),
            ("Scalability", 0.20),
        ]),
        ("Price", [
            ("Base Pricing", 0.40),
            ("Volume Discounts", 0.30),
            ("Total Cost of Ownership", 0.30),
        ]),
        ("Service", [
            ("Customer Support Quality", 0.30),
            ("Response Time", 0.25),
            ("Onboarding Experience", 0.20),
            ("Documentation & Training", 0.25),
        ]),
        ("Brand", [
            ("Market Reputation", 0.35),
            ("Industry Recognition", 0.30),
            ("Customer Loyalty", 0.35),
        ]),
    ]

    # Raw scores (1-10) for each company across all 15 criteria
    # Order: NovaTech, Apex, Pinnacle, Meridian, Vanguard, Summit, Catalyst
    scores = [
        # Product
        [8, 7, 9, 6, 8, 7, 5],   # Feature Completeness
        [9, 6, 7, 8, 7, 5, 6],   # Ease of Use
        [8, 8, 8, 7, 6, 7, 7],   # Reliability
        [7, 5, 8, 9, 7, 4, 6],   # Innovation
        [8, 7, 7, 6, 9, 6, 5],   # Scalability
        # Price
        [6, 8, 5, 7, 6, 9, 7],   # Base Pricing
        [7, 7, 6, 6, 5, 8, 8],   # Volume Discounts
        [6, 7, 5, 7, 6, 8, 7],   # Total Cost of Ownership
        # Service
        [9, 6, 7, 7, 8, 5, 6],   # Customer Support Quality
        [8, 5, 7, 6, 8, 6, 7],   # Response Time
        [7, 6, 8, 5, 7, 6, 5],   # Onboarding Experience
        [8, 5, 7, 6, 7, 5, 6],   # Documentation & Training
        # Brand
        [7, 8, 9, 6, 7, 6, 5],   # Market Reputation
        [6, 7, 8, 5, 6, 5, 4],   # Industry Recognition
        [8, 7, 8, 6, 7, 5, 6],   # Customer Loyalty
    ]

    # --- Styling ---
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    cat_font = Font(name="Calibri", size=11, bold=True, color="2F5496")
    cat_fill = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )

    # --- Row 1: Title ---
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "Competitive Analysis Matrix"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="1F3864")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # --- Row 2: Column Headers ---
    headers = ["Criteria", "Weight"] + companies
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Set column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 10
    for col_letter in ["C", "D", "E", "F", "G", "H", "I"]:
        ws.column_dimensions[col_letter].width = 18

    # --- Data rows ---
    current_row = 3
    score_idx = 0

    for cat_name, criteria_list in categories:
        # Category header row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
        cat_cell = ws.cell(row=current_row, column=1, value=f"Category: {cat_name}")
        cat_cell.font = cat_font
        cat_cell.fill = cat_fill
        cat_cell.alignment = Alignment(horizontal="left", vertical="center")
        for col in range(1, 10):
            ws.cell(row=current_row, column=col).fill = cat_fill
            ws.cell(row=current_row, column=col).border = thin_border
        current_row += 1

        # Criteria rows
        for crit_name, weight in criteria_list:
            ws.cell(row=current_row, column=1, value=crit_name).border = thin_border
            weight_cell = ws.cell(row=current_row, column=2, value=weight)
            weight_cell.number_format = '0.00'
            weight_cell.alignment = center_align
            weight_cell.border = thin_border

            for comp_idx in range(7):
                score_cell = ws.cell(row=current_row, column=comp_idx + 3, value=scores[score_idx][comp_idx])
                score_cell.alignment = center_align
                score_cell.border = thin_border

            score_idx += 1
            current_row += 1

    # --- Separator row ---
    separator_row = current_row
    current_row += 1

    # --- Summary Section ---
    # Category sub-score labels
    summary_start = current_row
    summary_labels = [
        "Product Sub-score",
        "Price Sub-score",
        "Service Sub-score",
        "Brand Sub-score",
        "",
        "Overall Weighted Score",
        "Ranking",
    ]

    for label in summary_labels:
        cell = ws.cell(row=current_row, column=1, value=label)
        if label and label != "":
            cell.font = Font(name="Calibri", size=11, bold=True)
            cell.border = thin_border
            for col in range(2, 10):
                ws.cell(row=current_row, column=col).border = thin_border
                ws.cell(row=current_row, column=col).alignment = center_align
        current_row += 1

    # --- SWOT Section ---
    current_row += 1  # blank row
    swot_row = current_row
    ws.merge_cells(start_row=swot_row, start_column=1, end_row=swot_row, end_column=9)
    swot_title = ws.cell(row=swot_row, column=1, value="SWOT-Style Summary")
    swot_title.font = Font(name="Calibri", size=14, bold=True, color="1F3864")
    swot_title.alignment = Alignment(horizontal="center")
    current_row += 1

    swot_categories = ["Strengths", "Weaknesses", "Opportunities", "Threats"]
    for swot_cat in swot_categories:
        cell = ws.cell(row=current_row, column=1, value=swot_cat)
        cell.font = Font(name="Calibri", size=11, bold=True, color="2F5496")
        # Leave columns B-I empty for SWOT text entries
        current_row += 1

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
