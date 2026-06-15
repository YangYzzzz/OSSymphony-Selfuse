"""
Reward Script: Competitive Analysis Matrix
Task ID: calc_wf_061
Domain: libreoffice_calc
Scoring:
  Component 1: SUMPRODUCT formulas for category sub-scores (0.25 pts)
  Component 2: SUM formulas for overall weighted score (0.15 pts)
  Component 3: RANK formulas for rankings (0.10 pts)
  Component 4: Radar chart with 7 series (0.20 pts)
  Component 5: Color scale conditional formatting on score cells (0.15 pts)
  Component 6: SWOT summary section with text entries (0.15 pts)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_061'


def persist_app_state(domain):
    """Try to save any open LibreOffice document."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Analysis' not in wb.sheetnames:
        print("CRITICAL: 'Analysis' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Analysis']

    # Component 1: SUMPRODUCT formulas for category sub-scores (0.25 points)
    # Golden has SUMPRODUCT formulas in C23:I26 for 4 categories
    # Initial has these cells empty
    try:
        sumproduct_count = 0
        expected_formulas = {
            23: ('$B$4:$B$8', '4:C8'),    # Product: weights B4:B8, scores row 4-8
            24: ('$B$10:$B$12', '10:C12'), # Price: weights B10:B12, scores row 10-12
            25: ('$B$14:$B$17', '14:C17'), # Service: weights B14:B17, scores row 14-17
            26: ('$B$19:$B$21', '19:C21'), # Brand: weights B19:B21, scores row 19-21
        }
        for row_num in [23, 24, 25, 26]:
            for col in range(3, 10):  # C through I
                cell_val = ws.cell(row=row_num, column=col).value
                if cell_val and isinstance(cell_val, str) and 'SUMPRODUCT' in cell_val.upper():
                    sumproduct_count += 1

        # 4 rows x 7 columns = 28 SUMPRODUCT formulas expected
        if sumproduct_count >= 28:
            print(f"PASS: Component 1 — All 28 SUMPRODUCT formulas found ({sumproduct_count}) (0.25 pts)")
            total_score += 0.25
        elif sumproduct_count >= 14:
            partial = 0.15
            print(f"PARTIAL: Component 1 — {sumproduct_count}/28 SUMPRODUCT formulas found ({partial} pts)")
            total_score += partial
        elif sumproduct_count >= 7:
            partial = 0.08
            print(f"PARTIAL: Component 1 — {sumproduct_count}/28 SUMPRODUCT formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {sumproduct_count}/28 SUMPRODUCT formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: SUM formulas for overall weighted score in row 28 (0.15 points)
    # Golden has =SUM(C23:C26) etc. in C28:I28
    # Initial has row 28 empty
    try:
        sum_count = 0
        for col in range(3, 10):  # C through I
            cell_val = ws.cell(row=28, column=col).value
            if cell_val and isinstance(cell_val, str):
                val_upper = cell_val.upper().replace(' ', '')
                if '=SUM(' in val_upper and '23' in val_upper and '26' in val_upper:
                    sum_count += 1

        if sum_count >= 7:
            print(f"PASS: Component 2 — All 7 SUM formulas in row 28 found (0.15 pts)")
            total_score += 0.15
        elif sum_count >= 3:
            partial = 0.08
            print(f"PARTIAL: Component 2 — {sum_count}/7 SUM formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {sum_count}/7 SUM formulas found in row 28")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: RANK formulas in row 29 (0.10 points)
    # Golden has =RANK(C28,$C$28:$I$28,0) etc.
    # Initial has row 29 empty
    try:
        rank_count = 0
        for col in range(3, 10):  # C through I
            cell_val = ws.cell(row=29, column=col).value
            if cell_val and isinstance(cell_val, str) and 'RANK' in cell_val.upper():
                rank_count += 1

        if rank_count >= 7:
            print(f"PASS: Component 3 — All 7 RANK formulas found in row 29 (0.10 pts)")
            total_score += 0.10
        elif rank_count >= 3:
            partial = 0.05
            print(f"PARTIAL: Component 3 — {rank_count}/7 RANK formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {rank_count}/7 RANK formulas found in row 29")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Radar chart with 7 series (0.20 points)
    # Golden has 1 RadarChart with 7 series; initial has 0 charts
    try:
        charts = ws._charts
        radar_charts = [ch for ch in charts if type(ch).__name__ == 'RadarChart']

        if len(radar_charts) >= 1:
            rc = radar_charts[0]
            series_count = len(rc.series)
            if series_count >= 7:
                print(f"PASS: Component 4 — RadarChart found with {series_count} series (0.20 pts)")
                total_score += 0.20
            elif series_count >= 4:
                partial = 0.12
                print(f"PARTIAL: Component 4 — RadarChart has {series_count}/7 series ({partial} pts)")
                total_score += partial
            else:
                partial = 0.05
                print(f"PARTIAL: Component 4 — RadarChart has only {series_count} series ({partial} pts)")
                total_score += partial
        elif len(charts) >= 1:
            # Some chart exists but not radar
            print(f"PARTIAL: Component 4 — Chart exists but type is {type(charts[0]).__name__}, expected RadarChart (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 4 — No charts found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Color scale conditional formatting on score cells (0.15 points)
    # Golden has 4 colorScale rules covering C4:I8, C10:I12, C14:I17, C19:I21
    # Initial has 0 CF rules
    try:
        cf_list = list(ws.conditional_formatting)
        color_scale_count = 0
        for cf in cf_list:
            for rule in cf.rules:
                if rule.type == 'colorScale':
                    color_scale_count += 1

        if color_scale_count >= 4:
            print(f"PASS: Component 5 — {color_scale_count} colorScale rules found (0.15 pts)")
            total_score += 0.15
        elif color_scale_count >= 2:
            partial = 0.08
            print(f"PARTIAL: Component 5 — {color_scale_count}/4 colorScale rules found ({partial} pts)")
            total_score += partial
        elif color_scale_count >= 1:
            partial = 0.04
            print(f"PARTIAL: Component 5 — {color_scale_count}/4 colorScale rules found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — No colorScale conditional formatting found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: SWOT summary section with text (0.15 points)
    # Golden has text in B32:B35 for Strengths, Weaknesses, Opportunities, Threats
    # Initial has labels in A32:A35 but B32:B35 are empty
    try:
        swot_labels = {32: 'Strengths', 33: 'Weaknesses', 34: 'Opportunities', 35: 'Threats'}
        swot_text_count = 0
        for row_num, label in swot_labels.items():
            cell_val = ws.cell(row=row_num, column=2).value  # Column B
            if cell_val and isinstance(cell_val, str) and len(cell_val.strip()) >= 20:
                swot_text_count += 1
                print(f"  SWOT '{label}': has text ({len(cell_val)} chars)")
            else:
                print(f"  SWOT '{label}': empty or too short (value: {repr(cell_val)})")

        if swot_text_count >= 4:
            print(f"PASS: Component 6 — All 4 SWOT entries have text (0.15 pts)")
            total_score += 0.15
        elif swot_text_count >= 2:
            partial = 0.08
            print(f"PARTIAL: Component 6 — {swot_text_count}/4 SWOT entries have text ({partial} pts)")
            total_score += partial
        elif swot_text_count >= 1:
            partial = 0.04
            print(f"PARTIAL: Component 6 — {swot_text_count}/4 SWOT entries have text ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 6 — No SWOT summary text found")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
