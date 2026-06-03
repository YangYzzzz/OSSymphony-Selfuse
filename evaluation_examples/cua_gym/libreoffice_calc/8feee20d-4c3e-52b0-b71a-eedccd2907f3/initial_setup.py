"""
Initial Setup: Weekly campaign comparison spreadsheet (Campaign A vs B)
Task ID: calc_sales_marketing_campaign_compare_030
Domain: libreoffice_calc

Creates a spreadsheet with 12 weeks of campaign metrics for two campaigns.
CTR columns are intentionally empty (task is to add them).
No chart (task is to create one).
No data validation dropdown (task is to add one).
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_marketing_campaign_compare_030'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: CampaignAB ---
    ws = wb.active
    ws.title = 'CampaignAB'

    # Headers
    headers = [
        'Week',
        'A_Impressions', 'A_Clicks', 'A_Conversions', 'A_Spend', 'A_CTR',
        'B_Impressions', 'B_Clicks', 'B_Conversions', 'B_Spend', 'B_CTR'
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri')
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # 12 weeks of realistic campaign data
    # Campaign A: larger brand campaign, higher impressions
    # Campaign B: targeted campaign, lower impressions but better conversion
    # CTR columns (F and K) are intentionally left empty
    weekly_data = [
        # Week, A_Impr, A_Clicks, A_Conv, A_Spend,  B_Impr, B_Clicks, B_Conv, B_Spend
        ('Week 1',  48200, 1205, 96,  2800.50,  32100, 963, 115, 2100.00),
        ('Week 2',  51300, 1284, 109, 2975.25,  34500, 1104, 128, 2340.00),
        ('Week 3',  49800, 1146, 91,  2650.00,  33200, 996, 119, 2175.50),
        ('Week 4',  53600, 1394, 125, 3120.75,  36800, 1178, 147, 2580.25),
        ('Week 5',  55100, 1378, 116, 3050.00,  38200, 1261, 155, 2720.00),
        ('Week 6',  57400, 1493, 134, 3380.50,  39700, 1270, 161, 2890.50),
        ('Week 7',  54900, 1318, 108, 2990.25,  37600, 1165, 143, 2500.00),
        ('Week 8',  58700, 1585, 142, 3520.00,  41200, 1360, 172, 3050.75),
        ('Week 9',  60200, 1626, 155, 3680.50,  42800, 1456, 183, 3210.00),
        ('Week 10', 62500, 1688, 163, 3850.25,  44100, 1499, 191, 3380.50),
        ('Week 11', 64800, 1814, 178, 4100.00,  46500, 1627, 204, 3700.25),
        ('Week 12', 67200, 1882, 189, 4320.75,  48900, 1761, 223, 3950.00),
    ]

    for r, row_data in enumerate(weekly_data, 2):
        week, a_impr, a_clicks, a_conv, a_spend, b_impr, b_clicks, b_conv, b_spend = row_data
        ws.cell(row=r, column=1, value=week)
        ws.cell(row=r, column=2, value=a_impr)
        ws.cell(row=r, column=3, value=a_clicks)
        ws.cell(row=r, column=4, value=a_conv)
        ws.cell(row=r, column=5, value=a_spend)
        # Column F (A_CTR) intentionally empty
        ws.cell(row=r, column=7, value=b_impr)
        ws.cell(row=r, column=8, value=b_clicks)
        ws.cell(row=r, column=9, value=b_conv)
        ws.cell(row=r, column=10, value=b_spend)
        # Column K (B_CTR) intentionally empty

    # Apply currency format to spend columns
    for r in range(2, 14):
        ws.cell(row=r, column=5).number_format = '$#,##0.00'
        ws.cell(row=r, column=10).number_format = '$#,##0.00'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 10
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        ws.column_dimensions[col].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheet: CampaignAB')
    print('Rows: 12 weeks of data (rows 2-13)')
    print('Columns: Week, A_Impressions, A_Clicks, A_Conversions, A_Spend, A_CTR(empty), '
          'B_Impressions, B_Clicks, B_Conversions, B_Spend, B_CTR(empty)')
    print('Note: CTR columns (F, K) are empty - to be filled by agent')
    print('Note: No chart - to be created by agent')
    print('Note: No dropdown validation - to be created by agent')


create_initial()
