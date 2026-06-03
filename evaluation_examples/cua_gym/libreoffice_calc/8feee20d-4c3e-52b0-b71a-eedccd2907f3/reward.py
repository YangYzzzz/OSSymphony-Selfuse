"""
Reward Script: Marketing Campaign A vs B weekly comparison
Task ID: calc_sales_marketing_campaign_compare_030
Domain: libreoffice_calc
Scoring:
  - Component 1: Campaign A CTR formulas (F2:F13 = =Cr/Br) with % format — 0.30 pts
  - Component 2: Campaign B CTR formulas (K2:K13 = =Hr/Gr) with % format — 0.20 pts
  - Component 3: Line chart with 2 series and correct title — 0.30 pts
  - Component 4: Data validation dropdown in M1 with CTR/CPC/Conversion Rate — 0.20 pts
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_marketing_campaign_compare_030'


def get_chart_title_text(chart):
    """Extract full concatenated text from an openpyxl chart title object."""
    try:
        title_obj = chart.title
        if title_obj is None:
            return None
        # Try to navigate the rich text structure and collect ALL run text
        if hasattr(title_obj, 'tx') and title_obj.tx:
            tx = title_obj.tx
            if hasattr(tx, 'rich') and tx.rich:
                all_text = ''
                for para in tx.rich.p:
                    for run in para.r:
                        if hasattr(run, 't') and run.t:
                            all_text += run.t
                if all_text:
                    return all_text
            if hasattr(tx, 'strRef') and tx.strRef:
                return str(tx.strRef)
        # Fallback: title might be a string directly
        if isinstance(title_obj, str):
            return title_obj
    except Exception:
        pass
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'CampaignAB' not in wb.sheetnames:
        print("CRITICAL: 'CampaignAB' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['CampaignAB']

    # Component 1: Campaign A CTR formulas in F2:F13 (0.30 points)
    # Each cell should contain =C{row}/B{row}, with percentage number format
    try:
        ctr_a_formula_ok = True
        ctr_a_format_ok = True
        missing_rows = []
        wrong_format_rows = []

        for row in range(2, 14):
            f_val = ws.cell(row=row, column=6).value  # column F
            f_fmt = ws.cell(row=row, column=6).number_format
            expected_formula = f'=C{row}/B{row}'

            if f_val is None or str(f_val).strip() != expected_formula:
                ctr_a_formula_ok = False
                missing_rows.append(f'F{row}: found {repr(f_val)}, expected {repr(expected_formula)}')

            # Check for percentage format (e.g., '0.00%', '0%', etc.)
            if f_fmt is None or '%' not in str(f_fmt):
                ctr_a_format_ok = False
                wrong_format_rows.append(f'F{row}: format={repr(f_fmt)}')

        if ctr_a_formula_ok:
            print("PASS: Component 1a — Campaign A CTR formulas (F2:F13) all correct (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1a — Campaign A CTR formulas missing/wrong: {missing_rows[:3]}")

        if ctr_a_format_ok:
            print("PASS: Component 1b — Campaign A CTR cells have percentage format (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1b — Campaign A CTR format not percentage: {wrong_format_rows[:3]}")

    except Exception as e:
        print(f"ERROR: Component 1 (Campaign A CTR) — {e}")

    # Component 2: Campaign B CTR formulas in K2:K13 (0.20 points)
    # Each cell should contain =H{row}/G{row}, with percentage number format
    try:
        ctr_b_formula_ok = True
        ctr_b_format_ok = True
        missing_rows_b = []
        wrong_format_rows_b = []

        for row in range(2, 14):
            k_val = ws.cell(row=row, column=11).value  # column K
            k_fmt = ws.cell(row=row, column=11).number_format
            expected_formula = f'=H{row}/G{row}'

            if k_val is None or str(k_val).strip() != expected_formula:
                ctr_b_formula_ok = False
                missing_rows_b.append(f'K{row}: found {repr(k_val)}, expected {repr(expected_formula)}')

            if k_fmt is None or '%' not in str(k_fmt):
                ctr_b_format_ok = False
                wrong_format_rows_b.append(f'K{row}: format={repr(k_fmt)}')

        if ctr_b_formula_ok and ctr_b_format_ok:
            print("PASS: Component 2 — Campaign B CTR formulas (K2:K13) correct with % format (0.20 pts)")
            total_score += 0.20
        elif ctr_b_formula_ok:
            print("PASS (partial): Component 2a — Campaign B CTR formulas present (0.10 pts)")
            total_score += 0.10
            print(f"FAIL: Component 2b — Campaign B CTR format not percentage: {wrong_format_rows_b[:3]}")
        else:
            print(f"FAIL: Component 2 — Campaign B CTR formulas missing/wrong: {missing_rows_b[:3]}")

    except Exception as e:
        print(f"ERROR: Component 2 (Campaign B CTR) — {e}")

    # Component 3: Line chart with 2 series and title "Weekly CTR: Campaign A vs B" (0.30 points)
    try:
        charts = ws._charts
        if len(charts) == 0:
            print("FAIL: Component 3 — No charts found on 'CampaignAB' sheet")
        else:
            # Find a line chart
            line_chart = None
            for chart in charts:
                if 'LineChart' in type(chart).__name__ or 'line' in str(type(chart).__name__).lower():
                    line_chart = chart
                    break

            if line_chart is None:
                print(f"FAIL: Component 3a — No line chart found. Found: {[type(c).__name__ for c in charts]}")
            else:
                print(f"PASS: Component 3a — Line chart found (0.10 pts)")
                total_score += 0.10

                # Check 2 series
                series_count = len(line_chart.series)
                if series_count >= 2:
                    print(f"PASS: Component 3b — Chart has {series_count} series (>= 2 required) (0.10 pts)")
                    total_score += 0.10
                else:
                    print(f"FAIL: Component 3b — Chart has {series_count} series, expected >= 2")

                # Check title "Weekly CTR: Campaign A vs B"
                # The title 'Weekly CTR: Campaign A vs B' contains Campaign A and 'vs B' (not literally 'Campaign B')
                title_text = get_chart_title_text(line_chart)
                title_str = str(title_text) if title_text else ''
                if title_text and 'Weekly CTR' in title_str and 'Campaign A' in title_str and ('vs B' in title_str or 'Campaign B' in title_str):
                    print(f"PASS: Component 3c — Chart title is 'Weekly CTR: Campaign A vs B': {repr(title_text)} (0.10 pts)")
                    total_score += 0.10
                elif title_text and 'CTR' in title_str and 'Campaign' in title_str:
                    print(f"PASS (partial): Component 3c — Chart title partially matches: {repr(title_text)} (0.05 pts)")
                    total_score += 0.05
                else:
                    print(f"FAIL: Component 3c — Chart title missing or wrong: {repr(title_text)}")

    except Exception as e:
        print(f"ERROR: Component 3 (Line Chart) — {e}")

    # Component 4: Data validation dropdown in M1 with CTR, CPC, Conversion Rate options (0.20 points)
    try:
        dvs = ws.data_validations.dataValidation
        if len(dvs) == 0:
            print("FAIL: Component 4 — No data validation found on sheet")
        else:
            found_valid_dv = False
            for dv in dvs:
                if dv.type == 'list' and dv.formula1:
                    formula_str = str(dv.formula1).strip('"')
                    options = [o.strip() for o in formula_str.split(',')]
                    has_ctr = any('CTR' in o for o in options)
                    has_cpc = any('CPC' in o for o in options)
                    has_conv = any('Conversion' in o or 'Conv' in o for o in options)

                    if has_ctr and has_cpc and has_conv:
                        found_valid_dv = True
                        # Also check if it applies to M1 or nearby
                        sqref_str = str(dv.sqref)
                        if 'M1' in sqref_str or 'M' in sqref_str:
                            print(f"PASS: Component 4 — Data validation dropdown in M1 with CTR/CPC/Conversion Rate (0.20 pts)")
                            total_score += 0.20
                        else:
                            print(f"PASS (partial): Component 4 — Data validation has correct options but applied to {sqref_str}, not M1 (0.10 pts)")
                            total_score += 0.10
                        break

            if not found_valid_dv:
                formulas = [str(dv.formula1) for dv in dvs]
                print(f"FAIL: Component 4 — Data validation found but missing CTR/CPC/Conversion Rate: {formulas}")

    except Exception as e:
        print(f"ERROR: Component 4 (Data Validation) — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
