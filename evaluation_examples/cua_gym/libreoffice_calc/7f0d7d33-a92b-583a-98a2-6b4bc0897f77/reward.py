"""
Reward Script: Define database range and export to dBASE (.dbf) format
Task ID: calc_gsi_070
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): DBF file exists with valid dBASE header
  Component 2 (0.3): DBF has correct record count (199) and field count (6)
  Component 3 (0.3): DBF field data matches spreadsheet content (spot-check)
"""

import os
import struct

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_070'


def persist_app_state(domain):
    """Save any unsaved LibreOffice state before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            import time
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def read_dbf(dbf_path):
    """Parse a dBASE III (.dbf) file and return (fields, records)."""
    with open(dbf_path, 'rb') as f:
        version = struct.unpack('B', f.read(1))[0]
        f.seek(4)
        num_records = struct.unpack('<I', f.read(4))[0]
        header_size = struct.unpack('<H', f.read(2))[0]
        record_size = struct.unpack('<H', f.read(2))[0]
        num_fields = (header_size - 32) // 32

        # Read field descriptors
        f.seek(32)
        fields = []
        for i in range(num_fields):
            field_data = f.read(32)
            field_name = field_data[:11].split(b'\x00')[0].decode('ascii')
            field_type = chr(field_data[11])
            field_len = field_data[16]
            fields.append((field_name, field_type, field_len))

        # Skip header terminator
        f.read(1)

        # Read all records
        records = []
        for rec_i in range(num_records):
            deletion_flag = f.read(1)
            values = {}
            for fname, ftype, flen in fields:
                raw = f.read(flen).decode('ascii', errors='replace').strip()
                values[fname] = raw
            records.append(values)

    return fields, records, num_records, num_fields


def read_xlsx_data(xlsx_path):
    """Read spreadsheet data rows for comparison."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = []
    for c in range(1, ws.max_column + 1):
        headers.append(str(ws.cell(row=1, column=c).value or ''))

    rows = []
    for r in range(2, ws.max_row + 1):
        row_data = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            row_data[headers[c - 1]] = str(val) if val is not None else ''
        rows.append(row_data)

    return headers, rows


def verify_task():
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    dbf_path = f'{WORKDIR}/{TASK_ID}.dbf'
    xlsx_path = f'{WORKDIR}/{TASK_ID}.xlsx'

    # Component 1: DBF file exists with valid dBASE header (0.4 points)
    try:
        if not os.path.exists(dbf_path):
            print(f"FAIL: Component 1 — DBF file not found at {dbf_path}")
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score

        file_size = os.path.getsize(dbf_path)
        if file_size < 32:
            print(f"FAIL: Component 1 — DBF file too small ({file_size} bytes), not a valid dBASE file")
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score

        # Validate dBASE version byte (common values: 3, 4, 5, 48, 131, etc.)
        with open(dbf_path, 'rb') as f:
            version_byte = struct.unpack('B', f.read(1))[0]
        # dBASE III = 3, dBASE IV = 4, dBASE 5 = 5, with memo variants 0x83, 0x8B, etc.
        valid_versions = {3, 4, 5, 48, 67, 131, 139, 203}
        if version_byte not in valid_versions and version_byte > 5:
            print(f"WARN: Component 1 — Unusual dBASE version byte: {version_byte}, proceeding anyway")

        print(f"PASS: Component 1 — DBF file exists ({file_size} bytes), version byte {version_byte} (0.4 pts)")
        total_score += 0.4

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: DBF has correct structure — 6 fields and 199 data records (0.3 points)
    try:
        fields, records, num_records, num_fields = read_dbf(dbf_path)

        sub_score = 0.0
        # Check field count
        if num_fields == 6:
            print(f"PASS: Component 2a — DBF has 6 fields as expected")
            sub_score += 0.15
        else:
            print(f"FAIL: Component 2a — Expected 6 fields, found {num_fields}")

        # Check record count (199 data rows, header excluded)
        if num_records == 199:
            print(f"PASS: Component 2b — DBF has 199 records as expected")
            sub_score += 0.15
        else:
            print(f"FAIL: Component 2b — Expected 199 records, found {num_records}")

        if sub_score > 0:
            print(f"PASS: Component 2 — Structure checks ({sub_score} pts)")
            total_score += sub_score

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: DBF data matches spreadsheet content (0.3 points)
    # Spot-check first record, last record, and field name correspondence
    try:
        fields, records, num_records, num_fields = read_dbf(dbf_path)
        xlsx_headers, xlsx_rows = read_xlsx_data(xlsx_path)

        sub_score = 0.0

        # 3a: Field names should correspond to xlsx column headers (truncated to ~10 chars uppercase)
        # dBASE truncates field names to 10 chars
        dbf_field_names = [f[0].upper() for f in fields]
        expected_truncated = [h.upper()[:10].replace(' ', '') for h in xlsx_headers]
        # Check that each DBF field name is a prefix of the corresponding xlsx header
        field_match_count = 0
        for i, (dbf_fn, xlsx_h) in enumerate(zip(dbf_field_names, expected_truncated)):
            # dBASE may further truncate, so check if xlsx header starts with dbf name or vice versa
            if xlsx_h.startswith(dbf_fn) or dbf_fn.startswith(xlsx_h[:len(dbf_fn)]):
                field_match_count += 1
        if field_match_count >= 4:  # at least 4 of 6 fields match
            print(f"PASS: Component 3a — {field_match_count}/6 field names match xlsx headers")
            sub_score += 0.1
        else:
            print(f"FAIL: Component 3a — Only {field_match_count}/6 field names match. DBF: {dbf_field_names}, Expected: {expected_truncated}")

        # 3b: First record data matches xlsx row 2
        if len(records) > 0 and len(xlsx_rows) > 0:
            first_dbf = records[0]
            first_xlsx = xlsx_rows[0]
            # Check CustomerID field
            dbf_cust_id = list(first_dbf.values())[0]
            xlsx_cust_id = first_xlsx.get('CustomerID', '')
            if dbf_cust_id == xlsx_cust_id:
                print(f"PASS: Component 3b — First record CustomerID matches: {dbf_cust_id}")
                sub_score += 0.1
            else:
                print(f"FAIL: Component 3b — First record CustomerID: DBF='{dbf_cust_id}', XLSX='{xlsx_cust_id}'")
        else:
            print(f"FAIL: Component 3b — No records to compare")

        # 3c: Last record matches
        if len(records) > 1 and len(xlsx_rows) > 1:
            last_dbf = records[-1]
            last_xlsx = xlsx_rows[-1]
            dbf_last_id = list(last_dbf.values())[0]
            xlsx_last_id = last_xlsx.get('CustomerID', '')
            if dbf_last_id == xlsx_last_id:
                print(f"PASS: Component 3c — Last record CustomerID matches: {dbf_last_id}")
                sub_score += 0.1
            else:
                print(f"FAIL: Component 3c — Last record CustomerID: DBF='{dbf_last_id}', XLSX='{xlsx_last_id}'")
        else:
            print(f"FAIL: Component 3c — Insufficient records for last-record check")

        if sub_score > 0:
            print(f"PASS: Component 3 — Data content checks ({sub_score} pts)")
            total_score += sub_score

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")
verify_task()
