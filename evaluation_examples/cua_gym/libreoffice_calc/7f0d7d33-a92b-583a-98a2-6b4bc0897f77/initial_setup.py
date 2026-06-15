"""
Initial Setup: Customer records spreadsheet for database range and dBASE export task
Task ID: calc_gsi_070
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_070'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"

    # Headers in row 1
    headers = ['CustomerID', 'CompanyName', 'ContactName', 'City', 'Phone', 'AccountBalance']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic data pools
    first_names = [
        'Sarah', 'Marcus', 'Elena', 'James', 'Priya', 'David', 'Mei', 'Carlos',
        'Anna', 'Robert', 'Yuki', 'Thomas', 'Fatima', 'Michael', 'Lena', 'Ahmed',
        'Jessica', 'Wei', 'Patricia', 'Dmitri', 'Olivia', 'Henrik', 'Amara', 'Felix',
        'Sophia', 'Kenji', 'Isabella', 'Omar', 'Natalie', 'Raj'
    ]
    last_names = [
        'Chen', 'Johnson', 'Petrov', 'Williams', 'Sharma', 'Kim', 'Mueller',
        'Garcia', 'Anderson', 'Tanaka', 'Brown', 'Hassan', 'Taylor', 'Wang',
        'Martinez', 'Singh', 'Johansson', 'Okafor', 'Thompson', 'Nakamura',
        'Davis', 'Ali', 'Wilson', 'Larsson', 'Patel'
    ]
    companies = [
        'Apex Manufacturing', 'Brightline Solutions', 'Cascade Industries',
        'Delta Logistics', 'Evergreen Supply Co', 'Falcon Electronics',
        'GlobalTech Systems', 'Horizon Medical', 'InnoVate Corp', 'JetStream Aviation',
        'Keystone Financial', 'Luminary Design', 'Meridian Consulting',
        'NorthStar Energy', 'Olympus Healthcare', 'Pinnacle Software',
        'Quantum Dynamics', 'Redwood Partners', 'Summit Trading', 'Trident Marine',
        'Unity Biotech', 'Vanguard Security', 'Westfield Properties',
        'Xenon Technologies', 'Yellowstone Mining', 'Zenith Aerospace',
        'Atlas Freight', 'BluePeak Analytics', 'CrownPoint Industries',
        'DawnBreaker Innovations'
    ]
    cities = [
        'New York', 'Los Angeles', 'Chicago', 'Houston', 'Phoenix', 'Philadelphia',
        'San Antonio', 'San Diego', 'Dallas', 'San Jose', 'Austin', 'Denver',
        'Seattle', 'Boston', 'Nashville', 'Portland', 'Atlanta', 'Miami',
        'Minneapolis', 'Detroit', 'Charlotte', 'Tampa', 'Orlando', 'Raleigh',
        'Salt Lake City'
    ]
    area_codes = ['212', '310', '312', '713', '602', '215', '210', '619',
                  '214', '408', '512', '303', '206', '617', '615', '503',
                  '404', '305', '612', '313']

    random.seed(42)  # reproducible data

    for row_idx in range(2, 201):  # rows 2..200 = 199 data rows
        cust_id = f'CUST-{row_idx - 1:04d}'
        company = random.choice(companies)
        first = random.choice(first_names)
        last = random.choice(last_names)
        contact = f'{first} {last}'
        city = random.choice(cities)
        area = random.choice(area_codes)
        phone = f'({area}) {random.randint(200, 999)}-{random.randint(1000, 9999)}'
        balance = round(random.uniform(500.0, 150000.0), 2)

        ws.cell(row=row_idx, column=1, value=cust_id)
        ws.cell(row=row_idx, column=2, value=company)
        ws.cell(row=row_idx, column=3, value=contact)
        ws.cell(row=row_idx, column=4, value=city)
        ws.cell(row=row_idx, column=5, value=phone)
        ws.cell(row=row_idx, column=6, value=balance)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 16

    # Number format for balance column
    for row_idx in range(2, 201):
        ws.cell(row=row_idx, column=6).number_format = '$#,##0.00'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
