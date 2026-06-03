"""
Initial Setup: HR Workbook with department lookup table and employee data
Task ID: calc_hr_named_range_lookup_021
Domain: libreoffice_calc

Creates a workbook with:
- 'Lookup Tables' sheet: dept codes A2:B8 (NO named ranges)
- 'Employees' sheet: 94 rows of employee data with VLOOKUP using raw range reference
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_named_range_lookup_021'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Lookup Tables ----
    ws_lookup = wb.active
    ws_lookup.title = 'Lookup Tables'

    # Headers
    ws_lookup['A1'] = 'Dept Code'
    ws_lookup['B1'] = 'Department Name'
    ws_lookup['A1'].font = Font(bold=True)
    ws_lookup['B1'].font = Font(bold=True)

    # Department code-to-name mappings (A2:B8 = 7 rows)
    dept_data = [
        ('ENG', 'Engineering'),
        ('MKT', 'Marketing'),
        ('FIN', 'Finance'),
        ('HR',  'Human Resources'),
        ('OPS', 'Operations'),
        ('SAL', 'Sales'),
        ('LEG', 'Legal'),
    ]
    for i, (code, name) in enumerate(dept_data, start=2):
        ws_lookup.cell(row=i, column=1, value=code)
        ws_lookup.cell(row=i, column=2, value=name)

    ws_lookup.column_dimensions['A'].width = 14
    ws_lookup.column_dimensions['B'].width = 22

    # NOTE: NO named range 'DeptCodes' is created here — that is the task

    # ---- Sheet 2: Employees ----
    ws_emp = wb.create_sheet('Employees')

    # Headers
    headers = ['Emp ID', 'Name', 'Dept Code', 'Department Name']
    for col, h in enumerate(headers, 1):
        cell = ws_emp.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    ws_emp.column_dimensions['A'].width = 10
    ws_emp.column_dimensions['B'].width = 22
    ws_emp.column_dimensions['C'].width = 12
    ws_emp.column_dimensions['D'].width = 22

    # Realistic employee data (94 employees, rows 2-95)
    dept_codes = ['ENG', 'MKT', 'FIN', 'HR', 'OPS', 'SAL', 'LEG']

    first_names = [
        'Sarah', 'Marcus', 'Emily', 'James', 'Linda', 'Robert', 'Olivia',
        'David', 'Sophia', 'Michael', 'Chloe', 'Daniel', 'Ava', 'Matthew',
        'Isabella', 'Andrew', 'Mia', 'Joshua', 'Charlotte', 'Ryan',
        'Amelia', 'Nathan', 'Harper', 'Tyler', 'Evelyn', 'Jacob', 'Abigail',
        'Logan', 'Ella', 'Dylan', 'Elizabeth', 'Ethan', 'Sofia', 'Brandon',
        'Avery', 'Justin', 'Scarlett', 'Benjamin', 'Grace', 'Samuel',
        'Zoe', 'Alexander', 'Riley', 'Nicholas', 'Aria', 'Christopher',
        'Lily', 'Anthony', 'Hannah', 'Kevin', 'Layla', 'Jonathan',
        'Nora', 'Brian', 'Leah', 'Eric', 'Aubrey', 'Stephen', 'Savannah',
        'Timothy', 'Brooklyn', 'Patrick', 'Bella', 'Gregory', 'Claire',
        'Scott', 'Skylar', 'Raymond', 'Lucy', 'Jeffrey', 'Paisley',
        'Frank', 'Everly', 'Dennis', 'Anna', 'Jerry', 'Caroline',
        'Walter', 'Genesis', 'Keith', 'Aaliyah', 'Arthur', 'Kennedy',
        'Lawrence', 'Ellie', 'Terry', 'Elena', 'Sean', 'Maya', 'Peter',
        'Naomi', 'Carl',
    ]

    last_names = [
        'Chen', 'Johnson', 'Williams', 'Brown', 'Davis', 'Miller', 'Wilson',
        'Moore', 'Taylor', 'Anderson', 'Thomas', 'Jackson', 'White', 'Harris',
        'Martin', 'Thompson', 'Garcia', 'Martinez', 'Robinson', 'Clark',
        'Rodriguez', 'Lewis', 'Lee', 'Walker', 'Hall', 'Allen', 'Young',
        'Hernandez', 'King', 'Wright', 'Lopez', 'Hill', 'Scott', 'Green',
        'Adams', 'Baker', 'Gonzalez', 'Nelson', 'Carter', 'Mitchell',
        'Perez', 'Roberts', 'Turner', 'Phillips', 'Campbell', 'Parker',
        'Evans', 'Edwards', 'Collins', 'Stewart', 'Sanchez', 'Morris',
        'Rogers', 'Reed', 'Cook', 'Morgan', 'Bell', 'Murphy', 'Bailey',
        'Rivera', 'Cooper', 'Richardson', 'Cox', 'Howard', 'Ward',
        'Torres', 'Peterson', 'Gray', 'Ramirez', 'James', 'Watson',
        'Brooks', 'Kelly', 'Sanders', 'Price', 'Bennett', 'Wood',
        'Barnes', 'Ross', 'Henderson', 'Coleman', 'Jenkins', 'Perry',
        'Powell', 'Long', 'Patterson', 'Hughes', 'Flores', 'Washington',
        'Butler', 'Simmons', 'Foster', 'Gonzales', 'Bryant',
    ]

    for i in range(94):
        row = i + 2
        emp_id = f'E{1001 + i:04d}'
        name = f'{first_names[i % len(first_names)]} {last_names[i % len(last_names)]}'
        dept_code = dept_codes[i % len(dept_codes)]

        ws_emp.cell(row=row, column=1, value=emp_id)
        ws_emp.cell(row=row, column=2, value=name)
        ws_emp.cell(row=row, column=3, value=dept_code)
        # VLOOKUP with raw range reference (NOT named range) — that's the initial state
        ws_emp.cell(row=row, column=4,
                    value=f"=VLOOKUP(C{row},'Lookup Tables'.$A:$B,2,0)")

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
