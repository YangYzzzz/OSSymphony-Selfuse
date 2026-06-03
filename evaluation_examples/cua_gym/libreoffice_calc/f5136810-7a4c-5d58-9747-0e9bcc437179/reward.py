"""
Reward Script: Create named range 'DeptCodes' and update VLOOKUP formulas
Task ID: calc_hr_named_range_lookup_021
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): Named range 'DeptCodes' defined with reference to 'Lookup Tables'.$A$2:$B$8
  Component 2 (0.5): All VLOOKUP formulas in Employees D2:D95 use 'DeptCodes' named range
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_hr_named_range_lookup_021'


def normalize_named_range_ref(ref_str):
    """
    Normalize a named range reference string for comparison.
    Strips quotes around sheet name and normalizes casing.
    E.g. "'Lookup Tables'!$A$2:$B$8" -> "lookup tables!$a$2:$b$8"
    """
    if ref_str is None:
        return ''
    # Remove surrounding quotes from sheet name portion
    normalized = ref_str.strip()
    # Remove single quotes around sheet name
    normalized = re.sub(r"'([^']+)'!", r"\1!", normalized)
    return normalized.lower()


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Verify both required sheets exist
    required_sheets = ['Lookup Tables', 'Employees']
    for sheet_name in required_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"CRITICAL: Required sheet '{sheet_name}' not found in workbook")
            print("REWARD: 0.0")
            return 0.0

    # Component 1: Named range 'DeptCodes' exists and references 'Lookup Tables'.$A$2:$B$8 (0.5 points)
    # This FAILS on initial file (no named ranges) and PASSES on golden file (DeptCodes defined)
    try:
        defined_names = wb.defined_names

        if 'DeptCodes' not in defined_names:
            print(f"FAIL: Component 1 — Named range 'DeptCodes' not found in workbook. "
                  f"Available defined names: {list(defined_names.keys())}")
        else:
            dr = defined_names['DeptCodes']
            ref_value = dr.attr_text if hasattr(dr, 'attr_text') else str(dr.value)

            # Normalize and check the reference points to Lookup Tables!$A$2:$B$8
            normalized_ref = normalize_named_range_ref(ref_value)
            # Expected: 'Lookup Tables'!$A$2:$B$8 (workbook scope or sheet scope)
            # After normalization: "lookup tables!$a$2:$b$8"
            expected_norm = "lookup tables!$a$2:$b$8"

            if expected_norm in normalized_ref:
                print(f"PASS: Component 1 — Named range 'DeptCodes' exists with correct reference '{ref_value}' (0.5 pts)")
                total_score += 0.5
            else:
                print(f"FAIL: Component 1 — Named range 'DeptCodes' exists but references '{ref_value}', expected reference to 'Lookup Tables'.$A$2:$B$8")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not check named ranges: {e}")

    # Component 2: All VLOOKUP formulas in Employees D2:D95 use 'DeptCodes' named range (0.5 points)
    # This FAILS on initial file (formulas use raw range 'Lookup Tables'.$A:$B) and PASSES on golden file
    try:
        ws_emp = wb['Employees']
        total_cells = 0
        cells_using_deptcodes = 0
        cells_using_old_ref = 0
        cells_with_other = 0
        sample_failures = []

        for row in range(2, 96):
            cell = ws_emp.cell(row=row, column=4)
            val = cell.value
            total_cells += 1

            if val is None:
                cells_with_other += 1
                if len(sample_failures) < 3:
                    sample_failures.append(f"D{row}: None (expected VLOOKUP formula)")
            elif isinstance(val, str) and val.upper().startswith('=VLOOKUP'):
                # Check if it uses DeptCodes named range (case-insensitive)
                if 'DeptCodes'.lower() in val.lower():
                    cells_using_deptcodes += 1
                elif 'Lookup Tables' in val or 'lookup tables' in val.lower():
                    cells_using_old_ref += 1
                    if len(sample_failures) < 3:
                        sample_failures.append(f"D{row}: {repr(val)} (still uses raw range)")
                else:
                    cells_with_other += 1
                    if len(sample_failures) < 3:
                        sample_failures.append(f"D{row}: {repr(val)} (unexpected formula)")
            else:
                cells_with_other += 1
                if len(sample_failures) < 3:
                    sample_failures.append(f"D{row}: {repr(val)} (not a VLOOKUP formula)")

        print(f"  Employees D2:D95 scan: {cells_using_deptcodes}/{total_cells} cells use DeptCodes, "
              f"{cells_using_old_ref} use old raw range, {cells_with_other} other")

        if cells_using_deptcodes == total_cells:
            print(f"PASS: Component 2 — All {total_cells} VLOOKUP formulas in D2:D95 use 'DeptCodes' named range (0.5 pts)")
            total_score += 0.5
        elif cells_using_deptcodes > 0 and cells_using_old_ref == 0:
            # Partial: some cells have DeptCodes but some have other issues
            print(f"FAIL: Component 2 — {cells_using_deptcodes}/{total_cells} cells use 'DeptCodes', "
                  f"but {cells_with_other} cells have unexpected content")
            if sample_failures:
                print(f"  Sample failures: {sample_failures}")
        else:
            print(f"FAIL: Component 2 — Only {cells_using_deptcodes}/{total_cells} cells use 'DeptCodes'. "
                  f"{cells_using_old_ref} cells still use raw range reference.")
            if sample_failures:
                print(f"  Sample failures: {sample_failures}")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not check VLOOKUP formulas in Employees: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
