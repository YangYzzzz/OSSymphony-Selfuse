"""
Initial Setup: Clone Formatting task - price list with specially formatted C3
Task ID: calc_gfl_043
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_043'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prices"

    # Row 1: Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "Bergmann Office Supplies - Product Price List 2025"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Row 2: Headers
    headers = ["Item", "Code", "Price", "Discount", "Net"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows 3-30 (28 items)
    items = [
        ["Premium Copy Paper A4 (500 sheets)", "OFS-1001", 24.99, 0.10, 22.49],
        ["Ballpoint Pens (12-pack, Blue)", "OFS-1002", 8.75, 0.05, 8.31],
        ["Heavy Duty Stapler", "OFS-1003", 35.50, 0.15, 30.18],
        ["Whiteboard Markers Set (8 colors)", "OFS-1004", 14.20, 0.08, 13.06],
        ["Manila File Folders (25-pack)", "OFS-1005", 18.60, 0.12, 16.37],
        ["Desk Organizer - Mesh Black", "OFS-1006", 42.30, 0.10, 38.07],
        ["Correction Tape (6-pack)", "OFS-1007", 11.95, 0.05, 11.35],
        ["Sticky Notes 3x3 (12 pads)", "OFS-1008", 9.80, 0.08, 9.02],
        ["Binder Clips Assorted (48-pack)", "OFS-1009", 7.45, 0.05, 7.08],
        ["Legal Pads Yellow (6-pack)", "OFS-1010", 15.30, 0.10, 13.77],
        ["Ergonomic Mouse Pad with Wrist Rest", "OFS-1011", 22.90, 0.12, 20.15],
        ["Push Pins Color Assortment (200ct)", "OFS-1012", 5.60, 0.05, 5.32],
        ["Laminating Pouches A4 (100-pack)", "OFS-1013", 29.75, 0.15, 25.29],
        ["Hanging File Folders (25-pack)", "OFS-1014", 21.40, 0.10, 19.26],
        ["Dry Erase Board 24x36 inch", "OFS-1015", 58.90, 0.18, 48.30],
        ["Paper Clips Jumbo (100ct)", "OFS-1016", 3.25, 0.05, 3.09],
        ["Desktop Calculator 12-Digit", "OFS-1017", 16.80, 0.08, 15.46],
        ["Rubber Bands Assorted (1 lb)", "OFS-1018", 6.95, 0.05, 6.60],
        ["Sheet Protectors Top-Load (100-pack)", "OFS-1019", 19.50, 0.10, 17.55],
        ["Tape Dispenser Heavy Duty", "OFS-1020", 12.40, 0.08, 11.41],
        ["Index Cards 3x5 Ruled (300ct)", "OFS-1021", 8.15, 0.05, 7.74],
        ["Clipboard with Storage", "OFS-1022", 13.70, 0.10, 12.33],
        ["Label Maker Tape Refill (3-pack)", "OFS-1023", 27.60, 0.12, 24.29],
        ["Pencil Sharpener Electric", "OFS-1024", 31.25, 0.15, 26.56],
        ["Envelope #10 White (100-pack)", "OFS-1025", 10.50, 0.08, 9.66],
        ["Presentation Binder 1-inch (4-pack)", "OFS-1026", 17.85, 0.10, 16.07],
        ["Scissors 8-inch Stainless Steel", "OFS-1027", 9.40, 0.05, 8.93],
        ["Desk Lamp LED Adjustable", "OFS-1028", 45.70, 0.18, 37.47],
    ]

    # Blue font color for the specially formatted C3
    blue_font_color = "FF0000FF"  # Blue

    for r, row_data in enumerate(items, 3):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            # Default styling for all data cells
            cell.font = Font(name="Calibri", size=11)

    # Specially format C3: currency, bold, blue font
    c3 = ws.cell(row=3, column=3)
    c3.font = Font(name="Calibri", size=11, bold=True, color="0000FF")
    c3.number_format = '$#,##0.00'

    # Discount column as percentage
    for r in range(3, 31):
        ws.cell(row=r, column=4).number_format = '0%'

    # Net column as plain number (no currency format)
    for r in range(3, 31):
        ws.cell(row=r, column=5).number_format = '#,##0.00'

    # Column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
