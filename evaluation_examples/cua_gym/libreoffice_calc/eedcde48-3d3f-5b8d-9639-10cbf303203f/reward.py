"""
Reward Script: Clone Formatting from C3 to C4:C30
Task ID: calc_gfl_043
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): Currency number format ($#,##0.00) applied to C4:C30
  Component 2 (0.35): Bold font applied to C4:C30
  Component 3 (0.20): Blue font color applied to C4:C30
  Component 4 (0.10): C3 formatting preserved (unchanged)
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_043'

# Expected values in C4:C30 (from initial file, must be preserved)
EXPECTED_VALUES = {
    4: 8.75, 5: 35.5, 6: 14.2, 7: 18.6, 8: 42.3, 9: 11.95,
    10: 9.8, 11: 7.45, 12: 15.3, 13: 22.9, 14: 5.6, 15: 29.75,
    16: 21.4, 17: 58.9, 18: 3.25, 19: 16.8, 20: 6.95, 21: 19.5,
    22: 12.4, 23: 8.15, 24: 13.7, 25: 27.6, 26: 31.25, 27: 10.5,
    28: 17.85, 29: 9.4, 30: 45.7
}


def persist_app_state(domain):
    """Attempt to save any unsaved LibreOffice edits via Ctrl+S."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify that Clone Formatting was applied from C3 to C4:C30.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify 'Prices' sheet exists
    if 'Prices' not in wb.sheetnames:
        print("CRITICAL: 'Prices' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Prices']

    # Read C3 reference formatting
    c3 = ws['C3']
    c3_nf = c3.number_format
    c3_bold = c3.font.bold
    try:
        c3_color = c3.font.color.rgb
    except Exception:
        c3_color = None

    # Component 1: Currency number format on C4:C30 (0.35 points)
    # In initial: all C4:C30 have 'General'. In golden: all have '$#,##0.00'.
    try:
        nf_match_count = 0
        total_cells = 27  # C4 through C30
        for r in range(4, 31):
            cell = ws.cell(row=r, column=3)
            if cell.number_format == '$#,##0.00':
                nf_match_count += 1

        nf_ratio = nf_match_count / total_cells
        if nf_ratio >= 0.95:
            print(f"PASS: Component 1 — Currency format on {nf_match_count}/{total_cells} cells (0.35 pts)")
            total_score += 0.35
        elif nf_ratio >= 0.5:
            partial = round(0.35 * nf_ratio, 2)
            print(f"PARTIAL: Component 1 — Currency format on {nf_match_count}/{total_cells} cells ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Currency format on only {nf_match_count}/{total_cells} cells")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Bold font on C4:C30 (0.35 points)
    # In initial: all C4:C30 have bold=False. In golden: all have bold=True.
    try:
        bold_match_count = 0
        for r in range(4, 31):
            cell = ws.cell(row=r, column=3)
            if cell.font.bold is True:
                bold_match_count += 1

        bold_ratio = bold_match_count / total_cells
        if bold_ratio >= 0.95:
            print(f"PASS: Component 2 — Bold font on {bold_match_count}/{total_cells} cells (0.35 pts)")
            total_score += 0.35
        elif bold_ratio >= 0.5:
            partial = round(0.35 * bold_ratio, 2)
            print(f"PARTIAL: Component 2 — Bold font on {bold_match_count}/{total_cells} cells ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Bold font on only {bold_match_count}/{total_cells} cells")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Blue font color on C4:C30 (0.20 points)
    # In initial: C4:C30 have no explicit font color. In golden: all have '000000FF' (blue).
    try:
        color_match_count = 0
        for r in range(4, 31):
            cell = ws.cell(row=r, column=3)
            try:
                cell_color = cell.font.color.rgb
                # Blue color: check that it ends with '0000FF' (ARGB format)
                if cell_color is not None and str(cell_color).upper().endswith('0000FF'):
                    color_match_count += 1
            except Exception:
                pass  # No color set

        color_ratio = color_match_count / total_cells
        if color_ratio >= 0.95:
            print(f"PASS: Component 3 — Blue font color on {color_match_count}/{total_cells} cells (0.20 pts)")
            total_score += 0.20
        elif color_ratio >= 0.5:
            partial = round(0.20 * color_ratio, 2)
            print(f"PARTIAL: Component 3 — Blue font color on {color_match_count}/{total_cells} cells ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Blue font color on only {color_match_count}/{total_cells} cells")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: C3 formatting preserved (0.10 points)
    # Verify C3 still has its original formatting: $#,##0.00, bold, blue color.
    # This component only awards points if C4:C30 have ALSO been changed (at least partial).
    # Otherwise it would pass on initial_env too.
    try:
        issues = []

        if c3_nf != '$#,##0.00':
            issues.append(f"number_format={c3_nf}")
        if c3_bold is not True:
            issues.append(f"bold={c3_bold}")
        if c3_color is None or not str(c3_color).upper().endswith('0000FF'):
            issues.append(f"color={c3_color}")

        c3_ok = len(issues) == 0

        # Only award points if some formatting was applied to C4:C30 (to avoid scoring on initial_env)
        formatting_applied = (nf_match_count + bold_match_count + color_match_count) > 0

        if c3_ok and formatting_applied:
            print(f"PASS: Component 4 — C3 formatting preserved (0.10 pts)")
            total_score += 0.10
        elif not c3_ok:
            print(f"FAIL: Component 4 — C3 formatting changed: {', '.join(issues)}")
        else:
            print(f"FAIL: Component 4 — No formatting applied to C4:C30 yet")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist any unsaved GUI state before verification
persist_app_state("libreoffice_calc")

# Run verification
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
