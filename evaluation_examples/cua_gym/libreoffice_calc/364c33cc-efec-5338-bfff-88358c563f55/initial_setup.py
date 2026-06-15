"""
Initial Setup: Create appointment spreadsheet without date validation
Task ID: calc_nrv_073
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_073'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Appointments"

    # --- Headers ---
    ws.cell(row=1, column=1, value="Appointment")
    ws.cell(row=1, column=2, value="Date")

    # Style headers
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    for col in range(1, 3):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Appointment data (realistic entries) ---
    # Column A: appointment descriptions, Column B: dates (all weekdays intentionally)
    appointments = [
        ("Annual Performance Review - Sarah Chen", "2025-03-10"),
        ("Client Presentation - Meridian Corp", "2025-03-12"),
        ("Budget Planning Meeting Q2", "2025-03-14"),
        ("Team Standup - Engineering Sprint 14", "2025-03-17"),
        ("Vendor Contract Renewal - TechFlow Inc", "2025-03-19"),
        ("HR Compliance Training Session", "2025-03-21"),
        ("Product Launch Strategy Review", "2025-03-24"),
        ("One-on-One with Marcus Johnson", "2025-03-26"),
        ("Cross-Department Sync - Marketing & Sales", "2025-03-28"),
        ("Quarterly Board Presentation Prep", "2025-03-31"),
        ("New Hire Orientation - April Cohort", "2025-04-02"),
        ("IT Security Audit Review", "2025-04-04"),
        ("Customer Feedback Analysis Session", "2025-04-07"),
        ("Office Renovation Planning", "2025-04-09"),
        ("End-of-Month Financial Closeout", "2025-04-11"),
    ]

    for r, (appt, date_str) in enumerate(appointments, 2):
        ws.cell(row=r, column=1, value=appt)
        ws.cell(row=r, column=2, value=date_str)

    # Set column widths for readability
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 15

    # NO data validation is applied - that's the task for the agent

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
