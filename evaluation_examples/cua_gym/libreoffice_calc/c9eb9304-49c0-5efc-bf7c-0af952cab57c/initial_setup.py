"""
Initial Setup: Art Exhibitions 2023 - Spreadsheet and Empty Document
Task ID: osworld_multi_apps_book_reading_rate_012
Domain: libreoffice_calc (multi-app: also creates a docx)
"""

import os
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from docx import Document

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_book_reading_rate_012'
XLSX_OUTPUT = f'{WORKDIR}/exhibitions_2023.xlsx'
DOCX_OUTPUT = f'{WORKDIR}/Desktop/lowest_attendance.docx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    # ----------------------------------------------------------------
    # 1. Create exhibitions_2023.xlsx
    # ----------------------------------------------------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exhibitions 2023"

    # Header row styling
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2E4057", end_color="FF2E4057", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Exhibition Name", "Museum", "Website", "Visitor Count"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border

    # Set column widths
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 16

    # Row height for headers
    ws.row_dimensions[1].height = 28

    # Exhibition data — Name, Museum, Website; Visitor Count is EMPTY (agent must fill it)
    data = [
        [
            "Vermeer",
            "Rijksmuseum",
            "https://www.rijksmuseum.nl/en/whats-on/exhibitions/vermeer",
            None,  # Visitor Count empty
        ],
        [
            "Picasso Celebration",
            "Musée National Picasso-Paris",
            "https://www.museepicassoparis.fr/en/picasso-celebration-1973-2023",
            None,
        ],
        [
            "Art and Climate Change",
            "Tate Modern",
            "https://www.tate.org.uk/whats-on/tate-modern/art-and-climate-change",
            None,
        ],
        [
            "After Impressionism",
            "National Gallery London",
            "https://www.nationalgallery.org.uk/exhibitions/after-impressionism",
            None,
        ],
        [
            "Manet/Degas",
            "Metropolitan Museum of Art",
            "https://www.metmuseum.org/exhibitions/manet-degas",
            None,
        ],
    ]

    data_font = Font(name="Calibri", size=11)
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    url_font = Font(name="Calibri", size=11, color="0000FF", underline="single")

    for r, row_data in enumerate(data, 2):
        ws.row_dimensions[r].height = 22
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = data_border
            cell.alignment = Alignment(vertical="center", wrap_text=(c == 3))
            if c == 3 and val:
                cell.font = url_font
            else:
                cell.font = data_font

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(XLSX_OUTPUT)
    print(f"Initial spreadsheet created: {XLSX_OUTPUT}")

    # ----------------------------------------------------------------
    # 2. Create empty lowest_attendance.docx on Desktop
    # ----------------------------------------------------------------
    os.makedirs(f"{WORKDIR}/Desktop", exist_ok=True)

    doc = Document()
    # Remove the default empty paragraph added by python-docx
    # Leave it as an empty document (the default paragraph is acceptable)
    doc.save(DOCX_OUTPUT)
    print(f"Empty document created: {DOCX_OUTPUT}")

    # ----------------------------------------------------------------
    # 3. GUI-ready startup: open both files
    # ----------------------------------------------------------------
    # Open the spreadsheet in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{XLSX_OUTPUT}"', delay_sec=2.0)
    # Open the empty document in LibreOffice Writer
    launch_gui(f'libreoffice --writer "{DOCX_OUTPUT}"', delay_sec=2.0)
    print("GUI_READY: launched LibreOffice Calc and Writer with DISPLAY=:0")


create_initial()
