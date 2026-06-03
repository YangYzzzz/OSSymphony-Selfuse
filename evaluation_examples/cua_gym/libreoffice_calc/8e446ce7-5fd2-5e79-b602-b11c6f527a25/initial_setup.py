"""
Initial Setup: Add text length validation to Description column in ProductCatalog
Task ID: calc_dop_validate_textlen_072
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_dop_validate_textlen_072'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ProductCatalog'

    # Headers: SKU (A), Product Name (B), Description (C), Category (D), Price (E)
    headers = ['SKU', 'Product Name', 'Description', 'Category', 'Price']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Product data — realistic entries with a mix of description lengths
    # Some entries are intentionally too short ('N/A', 'TBD', 'n')
    # Some entries are intentionally too long (300+ chars)
    data = [
        # SKU, Product Name, Description, Category, Price
        ('SKU-001', 'Ergonomic Office Chair',
         'High-back ergonomic office chair with lumbar support, adjustable armrests, and breathable mesh back for all-day comfort.',
         'Furniture', 349.99),
        ('SKU-002', 'Wireless Mechanical Keyboard',
         'N/A',
         'Electronics', 129.99),
        ('SKU-003', 'Ultra-Wide Monitor 34"',
         'Curved 34-inch ultra-wide IPS monitor with 3440x1440 resolution, 144Hz refresh rate, HDR support, and USB-C connectivity for professionals.',
         'Electronics', 699.00),
        ('SKU-004', 'Adjustable Standing Desk',
         'TBD',
         'Furniture', 549.00),
        ('SKU-005', 'Noise-Cancelling Headphones',
         'Over-ear Bluetooth headphones with active noise cancellation, 30-hour battery life, and premium audio drivers for immersive listening.',
         'Audio', 279.95),
        ('SKU-006', 'Portable SSD 1TB',
         'n',
         'Storage', 109.99),
        ('SKU-007', 'Webcam 4K',
         'Professional 4K webcam with autofocus, built-in noise-reducing microphone, and wide-angle lens for crystal-clear video conferencing and streaming.',
         'Electronics', 199.00),
        ('SKU-008', 'Desk Lamp LED',
         'A lamp.',
         'Lighting', 45.50),
        ('SKU-009', 'USB-C Hub 7-in-1',
         'Compact 7-in-1 USB-C hub with 4K HDMI, 100W PD charging, two USB-A 3.0 ports, SD and microSD card readers, and Gigabit Ethernet.',
         'Accessories', 59.99),
        ('SKU-010', 'Gaming Mouse Pro',
         'High-precision gaming mouse featuring a 25,600 DPI optical sensor, customizable RGB lighting, six programmable buttons, and a lightweight ergonomic design ideal for both competitive and casual gaming sessions.',
         'Peripherals', 79.99),
        ('SKU-011', 'Laptop Cooling Pad',
         'Dual-fan laptop cooling pad compatible with 15-17 inch laptops, featuring adjustable height, USB-powered fans, and a non-slip surface.',
         'Accessories', 34.99),
        ('SKU-012', 'Smart Power Strip',
         'OK',
         'Electronics', 49.99),
        ('SKU-013', 'Wireless Charging Pad',
         '15W fast wireless charging pad with Qi compatibility, anti-slip surface, LED indicator, and universal compatibility for all Qi-enabled smartphones.',
         'Accessories', 29.99),
        ('SKU-014', 'Portable Projector',
         'Mini portable projector with 1080p full HD resolution, built-in speakers, Wi-Fi and Bluetooth connectivity, and up to 3-hour battery life for on-the-go presentations and movie nights.',
         'Electronics', 349.00),
        ('SKU-015', 'Cable Management Kit',
         'Good product for cables.',
         'Accessories', 19.99),
        ('SKU-016', 'Monitor Stand Arm',
         'Fully adjustable dual monitor stand arm supporting screens up to 32 inches and 17.6 lbs each, with 360-degree rotation, cable management, and VESA 75x75 and 100x100 compatibility.',
         'Furniture', 89.00),
        ('SKU-017', 'Mechanical Pencil Set',
         'TBD',
         'Stationery', 14.99),
        ('SKU-018', 'Notebook A5 Hardcover',
         'Premium A5 hardcover notebook with 200 dot-grid pages, 100gsm acid-free paper, ribbon bookmark, elastic closure band, and inner pocket — ideal for journaling, sketching, and professional notes.',
         'Stationery', 22.50),
        ('SKU-019', 'Desk Organizer',
         'Bamboo desk organizer with multiple compartments for pens, papers, sticky notes, and small accessories, helping to keep your workspace tidy and efficient.',
         'Furniture', 38.00),
        ('SKU-020', 'Blue Light Glasses',
         'N/A',
         'Eyewear', 39.99),
        ('SKU-021', 'Compact Laser Printer',
         'Monochrome laser printer with 30ppm print speed, automatic duplex printing, 250-sheet paper tray, wireless connectivity, and compatibility with multiple OS platforms.',
         'Office Equipment', 229.00),
        ('SKU-022', 'Shredder Cross-Cut',
         'Cross-cut paper shredder capable of shredding 8 sheets at a time, with a 6-gallon bin, credit card shredding slot, and automatic jam-prevention technology.',
         'Office Equipment', 89.99),
        ('SKU-023', 'Electric Stapler',
         'Fast.',
         'Stationery', 29.99),
        ('SKU-024', 'Whiteboard 48x36',
         'Magnetic dry-erase whiteboard measuring 48x36 inches with aluminium frame, marker tray, and easy-mount hardware for home offices and conference rooms.',
         'Office Equipment', 79.00),
        ('SKU-025', 'Label Maker',
         'Handheld label maker with a QWERTY keyboard, 180 dpi print resolution, Bluetooth connectivity, and rechargeable battery — compatible with multiple tape widths and styles.',
         'Stationery', 44.99),
        ('SKU-026', 'Document Scanner',
         'This high-speed document scanner supports duplex scanning at 25 pages per minute, features an automatic document feeder with a 50-page capacity, offers 600 dpi resolution, has built-in Wi-Fi and USB connectivity, and is compatible with Windows and macOS operating systems; it also includes OCR software for converting scanned documents into searchable and editable PDFs for maximum productivity in any office setting.',
         'Office Equipment', 319.00),
        ('SKU-027', 'Filing Cabinet 3-Drawer',
         'Three-drawer vertical filing cabinet in steel construction with anti-tilt safety mechanism, lock and key security, smooth full-extension slides, and support for letter and legal-size hanging folders.',
         'Furniture', 159.00),
        ('SKU-028', 'Conference Phone',
         'n',
         'Audio', 249.00),
        ('SKU-029', 'Surge Protector 8-Outlet',
         'Eight-outlet surge protector with 2100 joule rating, two USB charging ports, 6-foot cord, and integrated circuit breaker for comprehensive device protection.',
         'Electronics', 39.99),
        ('SKU-030', 'Desk Calendar 2026',
         'Good.',
         'Stationery', 12.99),
        ('SKU-031', 'Ergonomic Mouse Vertical',
         'Vertical ergonomic mouse designed to reduce wrist strain with a natural handshake grip, 2400 DPI adjustable sensor, silent click buttons, and 18-month battery life.',
         'Peripherals', 49.99),
        ('SKU-032', 'Laptop Stand Adjustable',
         'Portable aluminum laptop stand with adjustable height and angle settings, foldable design for travel, heat dissipation vents, and compatibility with MacBook and most 10-17 inch notebooks.',
         'Accessories', 35.00),
        ('SKU-033', 'Ink Cartridge Set',
         'TBD',
         'Stationery', 24.99),
        ('SKU-034', 'Smart Desk Clock',
         'Multifunctional smart desk clock featuring LED display, temperature and humidity readings, USB charging port, dual alarm settings, and a 180-degree rotating stand.',
         'Electronics', 42.00),
        ('SKU-035', 'Cable Organizer Sleeve',
         'Flexible neoprene cable management sleeve in 6-foot length, zipper closure, compatible with cables up to 0.5-inch diameter, keeping your desk free of tangled wires.',
         'Accessories', 12.99),
        ('SKU-036', 'Anti-Fatigue Mat',
         'This premium anti-fatigue standing desk mat is engineered with a multi-layer foam core and a non-slip rubber base to provide exceptional cushioning and support for prolonged standing; its beveled edges prevent tripping, the stain-resistant surface is easy to clean, and its 3/4-inch thickness delivers lasting comfort for home offices and commercial environments where employees stand for extended periods throughout the day.',
         'Furniture', 89.95),
        ('SKU-037', 'Bluetooth Speaker Compact',
         'Pocket-sized Bluetooth 5.0 speaker delivering rich stereo sound, 12-hour battery life, IPX7 waterproof rating, and built-in microphone for hands-free calling.',
         'Audio', 59.00),
        ('SKU-038', 'Desk Pad XL',
         'N/A',
         'Accessories', 27.99),
        ('SKU-039', 'Video Conference Camera Bar',
         'All-in-one video conference camera bar with 4K lens, 120-degree field of view, AI-powered auto-framing, built-in speaker and microphone array, and USB-C and HDMI connections.',
         'Electronics', 499.00),
        ('SKU-040', 'Monitor Privacy Screen',
         'Reversible monitor privacy filter for 27-inch screens (16:9) with matte/gloss sides, 60-degree viewing angle limitation, blue light reduction, and anti-glare coating.',
         'Accessories', 44.99),
        ('SKU-041', 'Pen Holder Magnetic',
         'OK',
         'Stationery', 16.99),
        ('SKU-042', 'USB-A Hub 4-Port',
         'Compact 4-port USB-A 3.0 hub with SuperSpeed 5Gbps data transfer, individual power switches and LED indicators, and a 2-foot cable for flexible positioning.',
         'Accessories', 19.99),
        ('SKU-043', 'Document Holder',
         'Adjustable document holder that attaches to the side of your monitor with a flexible arm, supporting A4/letter-size papers and tablets for ergonomic reading.',
         'Accessories', 28.00),
        ('SKU-044', 'Heavy-Duty Tape Dispenser',
         'n',
         'Stationery', 11.99),
        ('SKU-045', 'Mini Fridge Office',
         'Compact 4-liter thermoelectric mini fridge suitable for desktop use, AC/DC powered, near-silent operation, and capable of cooling beverages to 32°F below ambient temperature.',
         'Appliances', 59.00),
        ('SKU-046', 'Coffee Mug Warmer',
         'Electric coffee mug warmer maintaining beverages at 131°F, with auto shut-off after 4 hours, non-slip rubber base, and compatibility with standard-sized mugs.',
         'Appliances', 25.99),
        ('SKU-047', 'Sticky Notes Assorted',
         'TBD',
         'Stationery', 9.99),
        ('SKU-048', 'Air Purifier Desktop',
         'Desktop HEPA air purifier covering up to 150 sq ft, with three-stage filtration, night mode, USB power, and a low-noise operation of under 35dB.',
         'Appliances', 79.00),
        ('SKU-049', 'Footrest Ergonomic',
         'Adjustable ergonomic footrest with massage surface and two height positions, helping to improve posture and reduce lower back strain during long sitting sessions.',
         'Furniture', 42.00),
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
