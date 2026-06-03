"""
Initial Setup: Freeze panes task - expense matrix spreadsheet
Task ID: calc_gfl_045
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_045'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Matrix'

    # --- Row 1: Merged title ---
    ws.merge_cells('A1:AD1')
    ws['A1'] = 'Annual Expense Matrix 2025'
    ws['A1'].font = Font(name='Arial', size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # --- Row 2: Column headers ---
    headers = [
        'Department',           # A
        'Jan', 'Feb', 'Mar',   # B-D
        'Apr', 'May', 'Jun',   # E-G
        'Jul', 'Aug', 'Sep',   # H-J
        'Oct', 'Nov', 'Dec',   # K-M
        'Q1 Total', 'Q2 Total', 'Q3 Total', 'Q4 Total',  # N-Q
        'Annual Total',         # R
        'Budget',               # S
        'Variance',             # T
        'Variance %',           # U
        'YoY Growth',           # V
        'Category',             # W
        'Cost Center',          # X
        'Approved By',          # Y
        'Last Updated',         # Z
        'Notes',                # AA
        'Priority',             # AB
        'Status',               # AC
        'Review Date',          # AD
    ]
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = header_border

    # --- Department names (rows 3-50) ---
    departments = [
        'Engineering', 'Marketing', 'Sales', 'Human Resources', 'Finance',
        'Legal', 'Operations', 'Customer Support', 'Research & Development',
        'Product Management', 'Quality Assurance', 'IT Infrastructure',
        'Data Analytics', 'Design', 'Supply Chain', 'Procurement',
        'Facilities', 'Training & Development', 'Public Relations',
        'Business Development', 'Compliance', 'Internal Audit',
        'Corporate Strategy', 'Investor Relations', 'Risk Management',
        'Environmental Safety', 'Logistics', 'Warehouse Operations',
        'Fleet Management', 'Security', 'Media & Communications',
        'Executive Office', 'Board Relations', 'Mergers & Acquisitions',
        'Tax Department', 'Treasury', 'Payroll', 'Benefits Administration',
        'Recruitment', 'Employee Relations', 'Organizational Development',
        'Innovation Lab', 'Digital Transformation', 'Cloud Services',
        'Network Operations', 'Application Development', 'Database Admin',
        'Project Management Office',
    ]

    categories = ['Fixed', 'Variable', 'Semi-Variable', 'Discretionary']
    cost_centers = ['CC-100', 'CC-200', 'CC-300', 'CC-400', 'CC-500',
                    'CC-600', 'CC-700', 'CC-800']
    approvers = ['Sarah Chen', 'Marcus Johnson', 'Elena Rodriguez',
                 'David Kim', 'Priya Patel', 'James O\'Brien']
    statuses = ['Active', 'Under Review', 'Approved', 'Pending']
    priorities = ['High', 'Medium', 'Low']

    random.seed(42)

    for row_idx, dept in enumerate(departments, 3):
        # Column A: Department name
        dept_cell = ws.cell(row=row_idx, column=1, value=dept)
        dept_cell.font = Font(name='Arial', size=10, bold=True)

        # Columns B-M: Monthly expenses (Jan-Dec)
        monthly = []
        base = random.randint(15000, 120000)
        for month_col in range(2, 14):
            val = round(base + random.uniform(-5000, 8000), 2)
            ws.cell(row=row_idx, column=month_col, value=val).number_format = '#,##0.00'
            monthly.append(val)

        # Columns N-Q: Quarterly totals
        q1 = round(sum(monthly[0:3]), 2)
        q2 = round(sum(monthly[3:6]), 2)
        q3 = round(sum(monthly[6:9]), 2)
        q4 = round(sum(monthly[9:12]), 2)
        ws.cell(row=row_idx, column=14, value=q1).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=15, value=q2).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=16, value=q3).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=17, value=q4).number_format = '#,##0.00'

        # Column R: Annual total
        annual = round(q1 + q2 + q3 + q4, 2)
        ws.cell(row=row_idx, column=18, value=annual).number_format = '#,##0.00'

        # Column S: Budget
        budget = round(annual * random.uniform(0.9, 1.15), 2)
        ws.cell(row=row_idx, column=19, value=budget).number_format = '#,##0.00'

        # Column T: Variance
        variance = round(budget - annual, 2)
        ws.cell(row=row_idx, column=20, value=variance).number_format = '#,##0.00'

        # Column U: Variance %
        var_pct = round(variance / budget * 100, 2) if budget != 0 else 0
        ws.cell(row=row_idx, column=21, value=var_pct).number_format = '0.00'

        # Column V: YoY Growth
        yoy = round(random.uniform(-8, 15), 2)
        ws.cell(row=row_idx, column=22, value=yoy).number_format = '0.00'

        # Column W: Category
        ws.cell(row=row_idx, column=23, value=random.choice(categories))

        # Column X: Cost Center
        ws.cell(row=row_idx, column=24, value=random.choice(cost_centers))

        # Column Y: Approved By
        ws.cell(row=row_idx, column=25, value=random.choice(approvers))

        # Column Z: Last Updated
        ws.cell(row=row_idx, column=26, value=f'2025-{random.randint(1,3):02d}-{random.randint(1,28):02d}')

        # Column AA: Notes
        ws.cell(row=row_idx, column=27, value='')

        # Column AB: Priority
        ws.cell(row=row_idx, column=28, value=random.choice(priorities))

        # Column AC: Status
        ws.cell(row=row_idx, column=29, value=random.choice(statuses))

        # Column AD: Review Date
        ws.cell(row=row_idx, column=30, value=f'2025-{random.randint(4,6):02d}-{random.randint(1,28):02d}')

    # Set column widths for readability
    ws.column_dimensions['A'].width = 28
    for col_letter in ['B','C','D','E','F','G','H','I','J','K','L','M',
                        'N','O','P','Q','R','S','T']:
        ws.column_dimensions[col_letter].width = 12

    # Row 2 height
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 25

    # IMPORTANT: No freeze panes set - that is the task
    ws.freeze_panes = None

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
