"""
Reward Script: Freeze first column and first two rows in expense_matrix.xlsx
Task ID: calc_gfl_045
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): freeze_panes is set to "B3" (freezes col A + rows 1-2)
  Component 2 (0.3): freeze is specifically at B3 (not some other cell)
                      AND the merged title row A1:AD1 is still intact
  Component 3 (0.2): data integrity — sample cells unchanged after freeze
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_045'


def persist_app_state(domain: str):
    """Save any unsaved LibreOffice state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Matrix' sheet must exist
    if 'Matrix' not in wb.sheetnames:
        print("CRITICAL: 'Matrix' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Matrix']

    # Component 1: freeze_panes is set (not None) and equals "B3" (0.5 points)
    # This is the core task requirement: freeze col A + rows 1-2 by setting freeze at B3
    try:
        fp = ws.freeze_panes
        if fp is not None and str(fp) == "B3":
            print(f"PASS: Component 1 — freeze_panes is '{fp}' (0.5 pts)")
            total_score += 0.5
        elif fp is not None:
            # Partial: something is frozen but not the right cell
            print(f"FAIL: Component 1 — freeze_panes is '{fp}', expected 'B3'")
        else:
            print(f"FAIL: Component 1 — freeze_panes is None, expected 'B3'")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: freeze at B3 AND merged title row A1:AD1 preserved (0.3 points)
    # Verifies that freeze was applied correctly AND the merge wasn't broken
    try:
        fp = ws.freeze_panes
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        title_merged = "A1:AD1" in merged_ranges

        if fp is not None and str(fp) == "B3" and title_merged:
            print(f"PASS: Component 2 — freeze at B3 AND merged title A1:AD1 intact (0.3 pts)")
            total_score += 0.3
        else:
            reasons = []
            if fp is None or str(fp) != "B3":
                reasons.append(f"freeze_panes={fp}")
            if not title_merged:
                reasons.append(f"merged title missing (found: {merged_ranges})")
            print(f"FAIL: Component 2 — {'; '.join(reasons)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: data integrity — key cells unchanged (0.2 points)
    # The task says "Data is unchanged". Verify sample cells match expected values.
    # These checks are COMPOUND: they only award points if freeze is also set correctly.
    try:
        fp = ws.freeze_panes
        if fp is None or str(fp) != "B3":
            print("FAIL: Component 3 — skipped because freeze not set correctly")
        else:
            issues_found = []

            # Check header row 2
            a2 = ws['A2'].value
            if a2 != 'Department':
                issues_found.append(f"A2={a2}, expected 'Department'")

            b2 = ws['B2'].value
            if b2 != 'Jan':
                issues_found.append(f"B2={b2}, expected 'Jan'")

            # Check a data cell
            a3 = ws['A3'].value
            if a3 != 'Engineering':
                issues_found.append(f"A3={a3}, expected 'Engineering'")

            # Check dimensions
            if ws.max_row < 50 or ws.max_column < 30:
                issues_found.append(f"dims={ws.max_row}x{ws.max_column}, expected >=50x30")

            if len(issues_found) == 0:
                print(f"PASS: Component 3 — data integrity verified (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 3 — data issues: {'; '.join(issues_found)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
