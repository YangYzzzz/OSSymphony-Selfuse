"""
Initial Setup: Student Loan Repayment Comparison Tool
Task ID: calc_wf_038
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_038'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    currency_fmt = '$#,##0.00'
    pct_fmt = '0.00%'
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    title_font = Font(name="Calibri", size=14, bold=True, color="2F5496")
    section_font = Font(name="Calibri", size=11, bold=True, color="2F5496")
    label_fill = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")

    def style_header_row(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def style_data_range(ws, start_row, end_row, cols):
        for r in range(start_row, end_row + 1):
            for c in range(1, cols + 1):
                ws.cell(row=r, column=c).border = thin_border

    # ============================================================
    # Sheet 1: Loan Details
    # ============================================================
    ws1 = wb.active
    ws1.title = "Loan Details"

    ws1["A1"] = "Student Loan Repayment Comparison Tool"
    ws1["A1"].font = Font(name="Calibri", size=16, bold=True, color="2F5496")
    ws1.merge_cells("A1:D1")

    ws1["A3"] = "Loan Parameters"
    ws1["A3"].font = title_font

    params = [
        ("Loan Balance", 45000, currency_fmt),
        ("Annual Interest Rate", 0.055, pct_fmt),
        ("Monthly Interest Rate", 0.055 / 12, '0.0000%'),
        ("Loan Origination Date", "2025-01-15", None),
        ("Borrower", "Jessica Martinez", None),
        ("Loan Servicer", "Federal Student Aid", None),
        ("Loan Type", "Direct Unsubsidized", None),
    ]
    for i, (label, val, fmt) in enumerate(params, 4):
        ws1.cell(row=i, column=1, value=label).font = section_font
        ws1.cell(row=i, column=1).fill = label_fill
        ws1.cell(row=i, column=1).border = thin_border
        c = ws1.cell(row=i, column=2, value=val)
        c.border = thin_border
        if fmt:
            c.number_format = fmt

    ws1["A13"] = "Repayment Plan Summary"
    ws1["A13"].font = title_font
    summary_headers = ["Plan", "Monthly Payment", "Total Paid", "Total Interest", "Term (Years)"]
    for c, h in enumerate(summary_headers, 1):
        ws1.cell(row=14, column=c, value=h)
    style_header_row(ws1, 14, 5)

    plan_names = ["Standard (10-Year)", "Graduated", "Extended (25-Year)"]
    for i, name in enumerate(plan_names, 15):
        ws1.cell(row=i, column=1, value=name).border = thin_border
        for c in range(2, 6):
            ws1.cell(row=i, column=c).border = thin_border
    # Leave columns B-E empty for formulas to be added by agent

    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 18
    ws1.column_dimensions["E"].width = 16

    # ============================================================
    # Sheet 2: Standard Plan (10-Year)
    # ============================================================
    ws2 = wb.create_sheet("Standard Plan")
    ws2["A1"] = "Standard 10-Year Repayment Plan"
    ws2["A1"].font = title_font
    ws2.merge_cells("A1:F1")

    ws2["A3"] = "Fixed Monthly Payment:"
    ws2["A3"].font = section_font
    # B3 left empty for PMT formula

    # Amortization table - yearly snapshots
    amort_headers = ["Year", "Beginning Balance", "Annual Payment", "Principal Paid", "Interest Paid", "Ending Balance"]
    for c, h in enumerate(amort_headers, 1):
        ws2.cell(row=5, column=c, value=h)
    style_header_row(ws2, 5, 6)

    # Pre-fill year numbers only; rest left for formulas
    principal = 45000
    rate = 0.055
    monthly_rate = rate / 12
    n_months = 120
    monthly_pmt = principal * (monthly_rate * (1 + monthly_rate)**n_months) / ((1 + monthly_rate)**n_months - 1)

    balance = principal
    for year in range(1, 11):
        row = 5 + year
        ws2.cell(row=row, column=1, value=year).border = thin_border
        ws2.cell(row=row, column=2).border = thin_border
        ws2.cell(row=row, column=3).border = thin_border
        ws2.cell(row=row, column=4).border = thin_border
        ws2.cell(row=row, column=5).border = thin_border
        ws2.cell(row=row, column=6).border = thin_border
        # Put beginning balance as raw data for context
        ws2.cell(row=row, column=2, value=round(balance, 2)).number_format = currency_fmt

        annual_interest = 0
        annual_principal = 0
        for m in range(12):
            interest = balance * monthly_rate
            princ = monthly_pmt - interest
            balance -= princ
            annual_interest += interest
            annual_principal += princ

        ws2.cell(row=row, column=3, value=round(monthly_pmt * 12, 2)).number_format = currency_fmt
        ws2.cell(row=row, column=4, value=round(annual_principal, 2)).number_format = currency_fmt
        ws2.cell(row=row, column=5, value=round(annual_interest, 2)).number_format = currency_fmt
        ws2.cell(row=row, column=6, value=round(max(balance, 0), 2)).number_format = currency_fmt

    # Totals row label
    ws2.cell(row=16, column=1, value="Total").font = Font(bold=True)
    ws2.cell(row=16, column=1).border = thin_border
    for c in range(2, 7):
        ws2.cell(row=16, column=c).border = thin_border

    for c_letter in ["A", "B", "C", "D", "E", "F"]:
        ws2.column_dimensions[c_letter].width = 20

    # ============================================================
    # Sheet 3: Graduated Plan
    # ============================================================
    ws3 = wb.create_sheet("Graduated Plan")
    ws3["A1"] = "Graduated Repayment Plan"
    ws3["A1"].font = title_font
    ws3.merge_cells("A1:F1")

    ws3["A3"] = "Payment Schedule (increases every 2 years):"
    ws3["A3"].font = section_font

    # Payment schedule table
    sched_headers = ["Period", "Years", "Monthly Payment"]
    for c, h in enumerate(sched_headers, 1):
        ws3.cell(row=4, column=c, value=h)
    style_header_row(ws3, 4, 3)

    # Graduated payments: start low, increase every 2 years over 10 years
    grad_payments = [
        ("Period 1", "Years 1-2", 350.00),
        ("Period 2", "Years 3-4", 420.00),
        ("Period 3", "Years 5-6", 510.00),
        ("Period 4", "Years 7-8", 620.00),
        ("Period 5", "Years 9-10", 755.00),
    ]
    for i, (period, years, pmt) in enumerate(grad_payments, 5):
        ws3.cell(row=i, column=1, value=period).border = thin_border
        ws3.cell(row=i, column=2, value=years).border = thin_border
        ws3.cell(row=i, column=3, value=pmt).number_format = currency_fmt
        ws3.cell(row=i, column=3).border = thin_border

    # Amortization
    ws3["A12"] = "Yearly Amortization Summary"
    ws3["A12"].font = section_font
    grad_amort_headers = ["Year", "Beginning Balance", "Monthly Payment", "Annual Payment", "Interest Paid", "Ending Balance"]
    for c, h in enumerate(grad_amort_headers, 1):
        ws3.cell(row=13, column=c, value=h)
    style_header_row(ws3, 13, 6)

    balance = 45000
    grad_monthly = [350, 350, 420, 420, 510, 510, 620, 620, 755, 755]
    for year in range(1, 11):
        row = 13 + year
        ws3.cell(row=row, column=1, value=year).border = thin_border
        ws3.cell(row=row, column=2, value=round(balance, 2)).number_format = currency_fmt
        ws3.cell(row=row, column=2).border = thin_border
        mp = grad_monthly[year - 1]
        ws3.cell(row=row, column=3, value=mp).number_format = currency_fmt
        ws3.cell(row=row, column=3).border = thin_border
        annual_pmt = mp * 12
        ws3.cell(row=row, column=4, value=annual_pmt).number_format = currency_fmt
        ws3.cell(row=row, column=4).border = thin_border
        annual_int = 0
        for m in range(12):
            interest = balance * monthly_rate
            princ_paid = mp - interest
            if princ_paid < 0:
                princ_paid = 0
                balance += (interest - mp)
            else:
                balance -= princ_paid
            annual_int += interest
        ws3.cell(row=row, column=5, value=round(annual_int, 2)).number_format = currency_fmt
        ws3.cell(row=row, column=5).border = thin_border
        ws3.cell(row=row, column=6, value=round(max(balance, 0), 2)).number_format = currency_fmt
        ws3.cell(row=row, column=6).border = thin_border

    ws3.cell(row=24, column=1, value="Total").font = Font(bold=True)
    ws3.cell(row=24, column=1).border = thin_border
    for c in range(2, 7):
        ws3.cell(row=24, column=c).border = thin_border

    for c_letter in ["A", "B", "C", "D", "E", "F"]:
        ws3.column_dimensions[c_letter].width = 20

    # ============================================================
    # Sheet 4: Extended Plan (25-Year)
    # ============================================================
    ws4 = wb.create_sheet("Extended Plan")
    ws4["A1"] = "Extended 25-Year Repayment Plan"
    ws4["A1"].font = title_font
    ws4.merge_cells("A1:F1")

    ws4["A3"] = "Fixed Monthly Payment:"
    ws4["A3"].font = section_font
    # B3 left empty for PMT formula

    ext_headers = ["Year", "Beginning Balance", "Annual Payment", "Principal Paid", "Interest Paid", "Ending Balance"]
    for c, h in enumerate(ext_headers, 1):
        ws4.cell(row=5, column=c, value=h)
    style_header_row(ws4, 5, 6)

    n_months_ext = 300
    monthly_pmt_ext = principal * (monthly_rate * (1 + monthly_rate)**n_months_ext) / ((1 + monthly_rate)**n_months_ext - 1)
    balance = principal
    for year in range(1, 26):
        row = 5 + year
        ws4.cell(row=row, column=1, value=year).border = thin_border
        ws4.cell(row=row, column=2, value=round(balance, 2)).number_format = currency_fmt
        ws4.cell(row=row, column=2).border = thin_border

        annual_interest = 0
        annual_principal = 0
        for m in range(12):
            interest = balance * monthly_rate
            princ = monthly_pmt_ext - interest
            balance -= princ
            annual_interest += interest
            annual_principal += princ

        ws4.cell(row=row, column=3, value=round(monthly_pmt_ext * 12, 2)).number_format = currency_fmt
        ws4.cell(row=row, column=3).border = thin_border
        ws4.cell(row=row, column=4, value=round(annual_principal, 2)).number_format = currency_fmt
        ws4.cell(row=row, column=4).border = thin_border
        ws4.cell(row=row, column=5, value=round(annual_interest, 2)).number_format = currency_fmt
        ws4.cell(row=row, column=5).border = thin_border
        ws4.cell(row=row, column=6, value=round(max(balance, 0), 2)).number_format = currency_fmt
        ws4.cell(row=row, column=6).border = thin_border

    ws4.cell(row=31, column=1, value="Total").font = Font(bold=True)
    ws4.cell(row=31, column=1).border = thin_border
    for c in range(2, 7):
        ws4.cell(row=31, column=c).border = thin_border

    for c_letter in ["A", "B", "C", "D", "E", "F"]:
        ws4.column_dimensions[c_letter].width = 20

    # ============================================================
    # Sheet 5: Prepayment Calculator
    # ============================================================
    ws5 = wb.create_sheet("Prepayment Calculator")
    ws5["A1"] = "Prepayment Calculator"
    ws5["A1"].font = Font(name="Calibri", size=16, bold=True, color="2F5496")
    ws5.merge_cells("A1:E1")

    ws5["A3"] = "How extra monthly payments affect your loan"
    ws5["A3"].font = section_font

    ws5["A5"] = "Base Plan:"
    ws5["A5"].font = section_font
    ws5["B5"] = "Standard (10-Year)"

    ws5["A6"] = "Base Monthly Payment:"
    ws5["A6"].font = section_font
    ws5.cell(row=6, column=2, value=round(monthly_pmt, 2)).number_format = currency_fmt

    ws5["A8"] = "Extra Monthly Payment:"
    ws5["A8"].font = Font(name="Calibri", size=12, bold=True, color="C00000")
    ws5["A8"].fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
    ws5.cell(row=8, column=2, value=200).number_format = currency_fmt
    ws5["B8"].fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")

    ws5["A10"] = "Prepayment Scenario Results"
    ws5["A10"].font = section_font

    prepay_headers = ["Extra Payment", "New Monthly Total", "Original Payoff (Months)",
                      "New Payoff (Months)", "Months Saved", "Interest Saved"]
    for c, h in enumerate(prepay_headers, 1):
        ws5.cell(row=11, column=c, value=h)
    style_header_row(ws5, 11, 6)

    # Pre-populate extra payment scenarios (raw data, no formulas)
    extras = [0, 50, 100, 150, 200, 300, 500]
    for i, extra in enumerate(extras, 12):
        ws5.cell(row=i, column=1, value=extra).number_format = currency_fmt
        ws5.cell(row=i, column=1).border = thin_border
        ws5.cell(row=i, column=2, value=round(monthly_pmt + extra, 2)).number_format = currency_fmt
        ws5.cell(row=i, column=2).border = thin_border
        ws5.cell(row=i, column=3, value=120).border = thin_border  # original 120 months

        # Iterative calc for new payoff
        bal = 45000
        months = 0
        total_int = 0
        total_pmt_with_extra = monthly_pmt + extra
        while bal > 0 and months < 360:
            interest = bal * monthly_rate
            total_int += interest
            princ = total_pmt_with_extra - interest
            if princ <= 0:
                break
            bal -= princ
            months += 1
            if bal <= 0:
                break
        ws5.cell(row=i, column=4, value=months).border = thin_border

        original_total_int = monthly_pmt * 120 - 45000
        int_saved = original_total_int - total_int
        ws5.cell(row=i, column=5, value=120 - months).border = thin_border
        ws5.cell(row=i, column=6, value=round(int_saved, 2)).number_format = currency_fmt
        ws5.cell(row=i, column=6).border = thin_border

    ws5.column_dimensions["A"].width = 22
    ws5.column_dimensions["B"].width = 22
    ws5.column_dimensions["C"].width = 24
    ws5.column_dimensions["D"].width = 22
    ws5.column_dimensions["E"].width = 18
    ws5.column_dimensions["F"].width = 18

    # ============================================================
    # Sheet 6: Comparison
    # ============================================================
    ws6 = wb.create_sheet("Comparison")
    ws6["A1"] = "Repayment Plan Comparison"
    ws6["A1"].font = Font(name="Calibri", size=16, bold=True, color="2F5496")
    ws6.merge_cells("A1:D1")

    comp_headers = ["Plan", "Total Paid", "Total Interest", "Monthly Payment Range"]
    for c, h in enumerate(comp_headers, 1):
        ws6.cell(row=3, column=c, value=h)
    style_header_row(ws6, 3, 4)

    # Raw comparison data (no formulas, no charts)
    comp_data = [
        ("Standard (10-Year)", round(monthly_pmt * 120, 2), round(monthly_pmt * 120 - 45000, 2), f"${monthly_pmt:.2f}"),
    ]
    # Calculate graduated totals
    grad_total_paid = sum(p * 12 * 2 for p in [350, 420, 510, 620, 755])
    comp_data.append(("Graduated", grad_total_paid, grad_total_paid - 45000, "$350 - $755"))
    comp_data.append(("Extended (25-Year)", round(monthly_pmt_ext * 300, 2), round(monthly_pmt_ext * 300 - 45000, 2), f"${monthly_pmt_ext:.2f}"))

    for i, (plan, total, interest, pmt_range) in enumerate(comp_data, 4):
        ws6.cell(row=i, column=1, value=plan).border = thin_border
        ws6.cell(row=i, column=2, value=total).number_format = currency_fmt
        ws6.cell(row=i, column=2).border = thin_border
        ws6.cell(row=i, column=3, value=interest).number_format = currency_fmt
        ws6.cell(row=i, column=3).border = thin_border
        ws6.cell(row=i, column=4, value=pmt_range).border = thin_border

    ws6.column_dimensions["A"].width = 24
    ws6.column_dimensions["B"].width = 18
    ws6.column_dimensions["C"].width = 18
    ws6.column_dimensions["D"].width = 24

    # NO charts in initial - agent must create them

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
