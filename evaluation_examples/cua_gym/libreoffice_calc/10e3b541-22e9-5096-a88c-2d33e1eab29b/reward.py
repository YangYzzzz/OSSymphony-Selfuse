"""
Reward Script: Student Loan Repayment Comparison Tool
Task ID: calc_wf_038
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20) - PMT formulas in Standard Plan B3 and Extended Plan B3
  Component 2 (0.20) - SUM totals formulas in Standard/Graduated/Extended plans
  Component 3 (0.20) - Loan Details summary rows 15-17 filled with values
  Component 4 (0.15) - Prepayment Calculator B6 uses PMT formula
  Component 5 (0.25) - Comparison sheet has a bar chart
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_038'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: PMT formulas for Standard and Extended plans (0.20 points)
    # Golden has =PMT(5.5%/12, 120, -45000) in Standard Plan B3
    # and =PMT(5.5%/12, 300, -45000) in Extended Plan B3
    # Initial has no formulas here (empty or hardcoded)
    try:
        comp1_score = 0.0
        std_ws = wb['Standard Plan']
        std_b3 = std_ws.cell(row=3, column=2).value
        if std_b3 and isinstance(std_b3, str) and 'PMT' in std_b3.upper():
            print(f"PASS: Standard Plan B3 has PMT formula: {std_b3}")
            comp1_score += 0.10
        else:
            print(f"FAIL: Standard Plan B3 expected PMT formula, found: {std_b3}")

        ext_ws = wb['Extended Plan']
        ext_b3 = ext_ws.cell(row=3, column=2).value
        if ext_b3 and isinstance(ext_b3, str) and 'PMT' in ext_b3.upper():
            print(f"PASS: Extended Plan B3 has PMT formula: {ext_b3}")
            comp1_score += 0.10
        else:
            print(f"FAIL: Extended Plan B3 expected PMT formula, found: {ext_b3}")

        total_score += comp1_score
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: SUM total formulas in amortization sheets (0.20 points)
    # Golden has SUM formulas in total rows of all three plan sheets
    # Initial has empty total rows
    try:
        comp2_score = 0.0
        sum_checks = 0

        # Standard Plan: C16, D16, E16 should have SUM formulas
        std_ws = wb['Standard Plan']
        for col in [3, 4, 5]:
            val = std_ws.cell(row=16, column=col).value
            if val and isinstance(val, str) and 'SUM' in val.upper():
                sum_checks += 1

        # Graduated Plan: D24, E24 should have SUM formulas
        grad_ws = wb['Graduated Plan']
        for col in [4, 5]:
            val = grad_ws.cell(row=24, column=col).value
            if val and isinstance(val, str) and 'SUM' in val.upper():
                sum_checks += 1

        # Extended Plan: C31, D31, E31 should have SUM formulas
        ext_ws = wb['Extended Plan']
        for col in [3, 4, 5]:
            val = ext_ws.cell(row=31, column=col).value
            if val and isinstance(val, str) and 'SUM' in val.upper():
                sum_checks += 1

        # 8 total SUM formula cells expected
        if sum_checks >= 6:
            comp2_score = 0.20
            print(f"PASS: Component 2 - {sum_checks}/8 SUM formulas found (0.20 pts)")
        elif sum_checks >= 3:
            comp2_score = 0.10
            print(f"PARTIAL: Component 2 - {sum_checks}/8 SUM formulas found (0.10 pts)")
        else:
            print(f"FAIL: Component 2 - only {sum_checks}/8 SUM formulas found")

        total_score += comp2_score
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Loan Details summary rows 15-17 filled (0.20 points)
    # Golden has actual values for Monthly Payment, Total Paid, Total Interest, Term
    # Initial has empty cells in B15:E17
    try:
        comp3_score = 0.0
        ld_ws = wb['Loan Details']

        filled_count = 0
        # Check B15 (Standard monthly payment ~488.37)
        b15 = ld_ws.cell(row=15, column=2).value
        if b15 is not None:
            filled_count += 1

        # Check C15 (Standard total paid ~58604)
        c15 = ld_ws.cell(row=15, column=3).value
        if c15 is not None:
            try:
                if abs(float(c15) - 58604.19) < 100:
                    filled_count += 1
            except (ValueError, TypeError):
                pass

        # Check D15 (Standard total interest ~13604)
        d15 = ld_ws.cell(row=15, column=4).value
        if d15 is not None:
            try:
                if abs(float(d15) - 13604.19) < 100:
                    filled_count += 1
            except (ValueError, TypeError):
                pass

        # Check C16 (Graduated total paid ~63720)
        c16 = ld_ws.cell(row=16, column=3).value
        if c16 is not None:
            try:
                if abs(float(c16) - 63720) < 500:
                    filled_count += 1
            except (ValueError, TypeError):
                pass

        # Check D16 (Graduated total interest ~18720)
        d16 = ld_ws.cell(row=16, column=4).value
        if d16 is not None:
            try:
                if abs(float(d16) - 18720) < 500:
                    filled_count += 1
            except (ValueError, TypeError):
                pass

        # Check D17 (Extended total interest ~37901)
        d17 = ld_ws.cell(row=17, column=4).value
        if d17 is not None:
            try:
                if abs(float(d17) - 37901.81) < 500:
                    filled_count += 1
            except (ValueError, TypeError):
                pass

        if filled_count >= 5:
            comp3_score = 0.20
            print(f"PASS: Component 3 - Loan Details summary has {filled_count}/6 correct values (0.20 pts)")
        elif filled_count >= 3:
            comp3_score = 0.10
            print(f"PARTIAL: Component 3 - Loan Details summary has {filled_count}/6 correct values (0.10 pts)")
        else:
            print(f"FAIL: Component 3 - Loan Details summary has {filled_count}/6 correct values")

        total_score += comp3_score
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Prepayment Calculator B6 uses PMT formula (0.15 points)
    # Golden has =PMT(5.5%/12, 120, -45000); Initial has hardcoded 488.37
    try:
        pp_ws = wb['Prepayment Calculator']
        pp_b6 = pp_ws.cell(row=6, column=2).value
        if pp_b6 and isinstance(pp_b6, str) and 'PMT' in pp_b6.upper():
            print(f"PASS: Component 4 - Prepayment B6 has PMT formula: {pp_b6} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 - Prepayment B6 expected PMT formula, found: {pp_b6}")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Comparison sheet has a bar chart (0.25 points)
    # Golden has 1 BarChart; Initial has 0 charts
    try:
        comp_ws = wb['Comparison']
        chart_count = len(comp_ws._charts)
        if chart_count >= 1:
            # Verify it's a bar chart type
            chart = comp_ws._charts[0]
            chart_class = chart.__class__.__name__
            if 'Bar' in chart_class:
                print(f"PASS: Component 5 - Comparison has {chart_count} bar chart(s) (0.25 pts)")
                total_score += 0.25
            else:
                # Still give partial credit for any chart
                print(f"PARTIAL: Component 5 - Comparison has chart but type is {chart_class}, expected Bar (0.15 pts)")
                total_score += 0.15
        else:
            print(f"FAIL: Component 5 - Comparison sheet has {chart_count} charts, expected >= 1")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
