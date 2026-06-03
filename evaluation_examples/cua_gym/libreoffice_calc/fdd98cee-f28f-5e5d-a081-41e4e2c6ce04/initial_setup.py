"""
Initial Setup: Column chart on 'Revenue' sheet with NO title
Task ID: calc_chart_title_add_016
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.chart import BarChart, Reference

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_title_add_016'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Revenue ---
    ws = wb.active
    ws.title = 'Revenue'

    # Headers
    ws['A1'] = 'Year'
    ws['B1'] = 'Revenue'

    # Data rows 2-7: years 2019-2024
    data = [
        (2019, 1200000),
        (2020, 980000),
        (2021, 1340000),
        (2022, 1580000),
        (2023, 1820000),
        (2024, 2100000),
    ]
    for r, (year, revenue) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=year)
        ws.cell(row=r, column=2, value=revenue)

    # --- Column chart using A1:B7 ---
    chart = BarChart()
    chart.type = 'col'   # vertical column chart
    # Intentionally NO title set — that is the task

    # Data: Revenue column including header row
    data_ref = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
    chart.add_data(data_ref, titles_from_data=True)

    # Categories: Year column (rows 2-7, excluding header)
    cats = Reference(ws, min_col=1, min_row=2, max_row=7)
    chart.set_categories(cats)

    # Place chart on sheet
    ws.add_chart(chart, 'D2')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
