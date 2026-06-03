"""
Reward Script: Set up research grant tracking sheet with formulas, warnings, conditional formatting, and pie chart.
Task ID: calc_edu_grant_tracker_023
Domain: libreoffice_calc
Scoring:
  - Component 1: Remaining formulas in E2:E11 (=C{n}-D{n} pattern) — 0.25 pts
  - Component 2: Pct Spent formulas in F2:F11 (=D{n}/C{n}) with percentage number format — 0.25 pts
  - Component 3: Warning IF formulas in G2:G11 (=IF(F{n}>0.8,"Review Required","")) — 0.20 pts
  - Component 4: Totals row formulas in E12 (SUM) and F12 (overall pct) — 0.10 pts
  - Component 5: Conditional formatting with yellow fill when Pct Spent > 80% — 0.10 pts
  - Component 6: Pie chart titled 'Grant Budget Allocation' with grant data — 0.10 pts
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_edu_grant_tracker_023'


def normalize_formula(f):
    """Normalize formula string for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the sheet exists (precondition gate)
    if 'Grants' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Grants' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Grants']

    # -----------------------------------------------------------------------
    # Component 1: Remaining formulas in E2:E11 (=C{n}-D{n} pattern) — 0.25 pts
    # Each data row should have a formula subtracting Spent from Total Amount.
    # FAILS on initial (E column is empty), PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        e_formula_correct = 0
        e_formula_total = 10
        for row in range(2, 12):
            val = ws.cell(row=row, column=5).value
            expected = f'=C{row}-D{row}'
            if val is not None and normalize_formula(val) == normalize_formula(expected):
                e_formula_correct += 1
            else:
                print(f"  FAIL Comp1: E{row} expected '{expected}', found {repr(val)}")

        if e_formula_correct == e_formula_total:
            print(f"PASS: Component 1 — All 10 Remaining formulas correct in E2:E11 (0.25 pts)")
            total_score += 0.25
        elif e_formula_correct >= 5:
            partial = round(0.25 * (e_formula_correct / e_formula_total), 4)
            print(f"PARTIAL: Component 1 — {e_formula_correct}/{e_formula_total} Remaining formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {e_formula_correct}/{e_formula_total} Remaining formulas (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Pct Spent formulas in F2:F11 (=D{n}/C{n}) + percentage format — 0.25 pts
    # Sub-component 2a: formula pattern correct (0.15 pts)
    # Sub-component 2b: number_format is percentage type (0.10 pts)
    # FAILS on initial (F column is empty), PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        f_formula_correct = 0
        f_format_correct = 0
        f_total = 10
        for row in range(2, 12):
            cell = ws.cell(row=row, column=6)
            val = cell.value
            expected = f'=D{row}/C{row}'
            if val is not None and normalize_formula(val) == normalize_formula(expected):
                f_formula_correct += 1
            else:
                print(f"  FAIL Comp2 formula: F{row} expected '{expected}', found {repr(val)}")
            # Check percentage format
            fmt = cell.number_format or ''
            if '%' in fmt:
                f_format_correct += 1

        if f_formula_correct == f_total:
            print(f"PASS: Component 2a — All 10 Pct Spent formulas correct in F2:F11 (0.15 pts)")
            total_score += 0.15
        elif f_formula_correct >= 5:
            partial = round(0.15 * (f_formula_correct / f_total), 4)
            print(f"PARTIAL: Component 2a — {f_formula_correct}/{f_total} Pct Spent formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2a — Only {f_formula_correct}/{f_total} Pct Spent formulas (0.0 pts)")

        if f_format_correct == f_total:
            print(f"PASS: Component 2b — All F2:F11 cells have percentage number format (0.10 pts)")
            total_score += 0.10
        elif f_format_correct >= 5:
            partial = round(0.10 * (f_format_correct / f_total), 4)
            print(f"PARTIAL: Component 2b — {f_format_correct}/{f_total} cells percentage format ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2b — Only {f_format_correct}/{f_total} cells have percentage format (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Warning IF formulas in G2:G11 — 0.20 pts
    # Checks for =IF(F{n}>0.8,"Review Required","") pattern.
    # FAILS on initial (G column is empty), PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        g_formula_correct = 0
        g_total = 10
        for row in range(2, 12):
            val = ws.cell(row=row, column=7).value
            expected = f'=IF(F{row}>0.8,"Review Required","")'
            if val is not None and normalize_formula(val) == normalize_formula(expected):
                g_formula_correct += 1
            else:
                print(f"  FAIL Comp3: G{row} expected '{expected}', found {repr(val)}")

        if g_formula_correct == g_total:
            print(f"PASS: Component 3 — All 10 Warning IF formulas correct in G2:G11 (0.20 pts)")
            total_score += 0.20
        elif g_formula_correct >= 5:
            partial = round(0.20 * (g_formula_correct / g_total), 4)
            print(f"PARTIAL: Component 3 — {g_formula_correct}/{g_total} Warning IF formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {g_formula_correct}/{g_total} Warning IF formulas (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: Totals row formulas in E12 and F12 — 0.10 pts
    # E12 should be =SUM(E2:E11); F12 should be an overall pct spent formula.
    # FAILS on initial (row 12 only has "Totals" label), PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        e12_val = ws.cell(row=12, column=5).value
        f12_val = ws.cell(row=12, column=6).value

        e12_matches = (e12_val is not None and
                       normalize_formula(e12_val) == normalize_formula('=SUM(E2:E11)'))
        if e12_matches:
            print(f"PASS: Component 4a — E12 has SUM formula: {repr(e12_val)} (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4a — E12 expected '=SUM(E2:E11)', found {repr(e12_val)}")

        # Accept =SUM(D2:D11)/SUM(C2:C11) or equivalent overall pct spent formula
        f12_matches = (f12_val is not None and
                       isinstance(f12_val, str) and
                       'SUM' in normalize_formula(f12_val) and
                       ('D' in normalize_formula(f12_val) or 'F' in normalize_formula(f12_val)))
        if f12_matches:
            print(f"PASS: Component 4b — F12 has overall pct formula: {repr(f12_val)} (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4b — F12 expected overall pct formula, found {repr(f12_val)}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -----------------------------------------------------------------------
    # Component 5: Conditional formatting (yellow fill when F > 0.8) — 0.10 pts
    # Checks that CF rules exist with threshold > 0.8 and a fill color.
    # FAILS on initial (no CF rules), PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        cf_rules = ws.conditional_formatting
        cf_rules_list = list(cf_rules)
        num_cf_ranges = len(cf_rules_list)

        if num_cf_ranges == 0:
            print(f"FAIL: Component 5 — No conditional formatting rules found (0.0 pts)")
        else:
            # Verify at least one rule references the F column > 0.8 threshold with a fill
            cf_valid_count = 0
            for cf_range in cf_rules_list:
                for rule in cf_range.rules:
                    formula_str = ''
                    if hasattr(rule, 'formula') and rule.formula:
                        formula_str = ' '.join(str(f) for f in rule.formula).upper().replace(' ', '')
                    has_threshold = '>0.8' in formula_str or ('$F' in formula_str and '0.8' in formula_str)
                    try:
                        has_fill = bool(rule.dxf and rule.dxf.fill and rule.dxf.fill.fgColor)
                    except Exception:
                        has_fill = False
                    if has_threshold and has_fill:
                        cf_valid_count += 1

            if cf_valid_count > 0:
                print(f"PASS: Component 5 — Conditional formatting with fill on rows where Pct Spent > 80% (0.10 pts)")
                total_score += 0.10
            elif num_cf_ranges > 0:
                print(f"PARTIAL: Component 5 — CF rules exist but no valid threshold+fill rule found (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 5 — CF rules do not match expected pattern (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # -----------------------------------------------------------------------
    # Component 6: Pie chart with title 'Grant Budget Allocation' — 0.10 pts
    # Sub-checks: chart exists (0.04), correct title (0.03), correct data refs (0.03).
    # FAILS on initial (no charts), PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        charts = ws._charts
        pie_charts = [c for c in charts if type(c).__name__ == 'PieChart']

        if len(pie_charts) == 0:
            print(f"FAIL: Component 6 — No pie chart found in Grants sheet (0.0 pts)")
        elif len(pie_charts) >= 1:
            print(f"PASS: Component 6 — Pie chart found in Grants sheet (0.04 pts)")
            total_score += 0.04

            chart = pie_charts[0]
            # Check title text
            title_text = ''
            try:
                if chart.title is not None:
                    t = chart.title
                    if isinstance(t, str):
                        title_text = t
                    else:
                        for para in t.tx.rich.p:
                            for run in para.r:
                                title_text += run.t
            except Exception:
                title_text = ''

            if 'grant budget allocation' in title_text.lower():
                print(f"PASS: Component 6a — Pie chart title is 'Grant Budget Allocation' (0.03 pts)")
                total_score += 0.03
            else:
                print(f"FAIL: Component 6a — Expected title 'Grant Budget Allocation', found: {repr(title_text)}")

            # Check data references use C column (Total Amount) and A column (Grant Names)
            data_ref_correct = False
            for s in chart.series:
                val_ref = ''
                cat_ref = ''
                try:
                    val_ref = s.val.numRef.f if s.val and s.val.numRef else ''
                except Exception:
                    pass
                try:
                    cat_ref = s.cat.numRef.f if s.cat and s.cat.numRef else ''
                except Exception:
                    pass
                val_uses_c = 'C' in val_ref.upper() and ('C2' in val_ref or '$C$2' in val_ref)
                cat_uses_a = 'A' in cat_ref.upper() and ('A2' in cat_ref or '$A$2' in cat_ref)
                if val_uses_c and cat_uses_a:
                    data_ref_correct = val_uses_c and cat_uses_a

            if data_ref_correct:
                print(f"PASS: Component 6b — Chart uses Total Amount (col C) and Grant Names (col A) (0.03 pts)")
                total_score += 0.03
            else:
                print(f"FAIL: Component 6b — Chart data refs do not match expected C/A columns")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score:.4f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
