"""
Initial Setup: Research grant tracking sheet with 10 grants
Task ID: calc_edu_grant_tracker_023
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_grant_tracker_023'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Grants ---
    ws = wb.active
    ws.title = 'Grants'

    # Headers in row 1
    headers = ['Grant Name', 'Funder', 'Total Amount', 'Spent', 'Remaining', 'Pct Spent', 'Warning']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # 10 grants data (realistic edu/research grant data)
    # Columns: Grant Name, Funder, Total Amount, Spent
    # E (Remaining), F (Pct Spent), G (Warning) are intentionally left EMPTY
    grants = [
        ('STEM Education Initiative',       'National Science Foundation',    250000.00,  198500.00),
        ('K-12 Literacy Improvement',        'Department of Education',        175000.00,   62300.00),
        ('Rural Schools Technology Grant',   'Gates Foundation',               320000.00,  289450.00),
        ('Early Childhood Learning Fund',    'Spencer Foundation',              95000.00,   41800.00),
        ('Teacher Professional Development', 'Wallace Foundation',             140000.00,  118900.00),
        ('STEM Diversity & Inclusion',        'NSF ADVANCE Program',           210000.00,   87600.00),
        ('College Access & Completion',      'Lumina Foundation',             185000.00,  160200.00),
        ('Digital Equity Initiative',        'Mozilla Foundation',              75000.00,   20500.00),
        ('Community College Partnership',    'Kresge Foundation',             300000.00,  256700.00),
        ('Research Methods Training',        'American Educational Research',   60000.00,   49100.00),
    ]

    for r, (name, funder, total, spent) in enumerate(grants, 2):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=funder)
        ws.cell(row=r, column=3, value=total)
        ws.cell(row=r, column=4, value=spent)
        # Columns E (5), F (6), G (7) intentionally left empty

        # Format currency columns
        ws.cell(row=r, column=3).number_format = '$#,##0.00'
        ws.cell(row=r, column=4).number_format = '$#,##0.00'

    # Row 12: Totals row (labels only, no formulas yet)
    ws.cell(row=12, column=1, value='Totals')
    ws.cell(row=12, column=1).font = Font(bold=True)
    # E12 and F12 are intentionally left empty

    # Column widths for readability
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

create_initial()
