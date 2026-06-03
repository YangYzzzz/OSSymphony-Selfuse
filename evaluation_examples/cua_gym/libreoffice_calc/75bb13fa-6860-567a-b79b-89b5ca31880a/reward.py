"""
Reward Script: Fitness Tracker Spreadsheet with Formulas, Chart, and Conditional Formatting
Task ID: calc_gen_personal_028
Domain: libreoffice_calc
Scoring:
  - Component 1: Weight change formulas in F3:F13 (0.25 pts)
  - Component 2: Trend indicator formulas in G3:G13 (0.25 pts)
  - Component 3: 4-week running average formulas in H6:H13 (0.20 pts)
  - Component 4: Line chart with 2 series and title 'Weight Progress' (0.20 pts)
  - Component 5: Conditional formatting on F3:F13 (green/red fills) (0.10 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_personal_028'
SHEET_NAME = 'FitnessTracker'


def normalize_formula(f):
    """Normalize formula for comparison: uppercase, no spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # -------------------------------------------------------------------------
    # Component 1: Weight change formulas in F3:F13 (0.25 points)
    # Each cell F(n) for n=3..13 should contain =C(n)-C(n-1)
    # This FAILS on initial (F3:F13 are empty) and PASSES on golden
    # -------------------------------------------------------------------------
    try:
        correct_f_formulas = 0
        total_f_expected = 11  # rows 3 through 13

        for row in range(3, 14):  # rows 3..13
            cell_val = ws.cell(row=row, column=6).value  # column F
            if cell_val is None:
                print(f"FAIL: Component 1 — F{row} is empty, expected =C{row}-C{row-1}")
                continue
            # Expected formula: =C(n)-C(n-1)
            expected = f'=C{row}-C{row - 1}'
            actual_norm = normalize_formula(str(cell_val))
            expected_norm = normalize_formula(expected)
            if actual_norm == expected_norm:
                correct_f_formulas += 1
            else:
                print(f"FAIL: Component 1 — F{row}: expected {expected!r}, found {cell_val!r}")

        if correct_f_formulas == total_f_expected:
            print(f"PASS: Component 1 — All {total_f_expected} weight change formulas correct in F3:F13 (0.25 pts)")
            total_score += 0.25
        elif correct_f_formulas >= 6:
            # Partial: more than half correct
            partial = round(0.25 * correct_f_formulas / total_f_expected, 4)
            print(f"PARTIAL: Component 1 — {correct_f_formulas}/{total_f_expected} weight change formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {correct_f_formulas}/{total_f_expected} weight change formulas correct (0 pts)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Trend indicator formulas in G3:G13 (0.25 points)
    # Each cell G(n) for n=3..13 should contain =IF(F(n)>0,"↑",IF(F(n)<0,"↓","→"))
    # This FAILS on initial (G3:G13 are empty) and PASSES on golden
    # -------------------------------------------------------------------------
    try:
        correct_g_formulas = 0
        total_g_expected = 11  # rows 3 through 13

        for row in range(3, 14):  # rows 3..13
            cell_val = ws.cell(row=row, column=7).value  # column G
            if cell_val is None:
                print(f"FAIL: Component 2 — G{row} is empty, expected trend IF formula")
                continue
            cell_str = str(cell_val)
            # Check that it's an IF formula referencing F(n) for trend
            # Accept flexible matching: must contain IF, F{row}, and the arrow characters
            cell_norm = normalize_formula(cell_str)
            # Must reference F(row) and contain IF and all three arrows
            expected_ref = f'F{row}'
            has_if = 'IF(' in cell_norm or 'IF(' in cell_str.upper()
            has_fref = expected_ref.upper() in cell_norm
            has_up = '↑' in cell_str
            has_down = '↓' in cell_str
            has_flat = '→' in cell_str

            if has_if and has_fref and has_up and has_down and has_flat:
                correct_g_formulas += 1
            else:
                print(f"FAIL: Component 2 — G{row}: formula validation failed. Value: {cell_val!r}")

        if correct_g_formulas == total_g_expected:
            print(f"PASS: Component 2 — All {total_g_expected} trend indicator formulas correct in G3:G13 (0.25 pts)")
            total_score += 0.25
        elif correct_g_formulas >= 6:
            partial = round(0.25 * correct_g_formulas / total_g_expected, 4)
            print(f"PARTIAL: Component 2 — {correct_g_formulas}/{total_g_expected} trend formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {correct_g_formulas}/{total_g_expected} trend formulas correct (0 pts)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: 4-week running average formulas in H6:H13 (0.20 points)
    # H6 = AVERAGE(C2:C5), H7 = AVERAGE(C3:C6), ... H13 = AVERAGE(C9:C12)
    # H2:H5 should be empty (insufficient history)
    # This FAILS on initial (H column is empty) and PASSES on golden
    # -------------------------------------------------------------------------
    try:
        correct_h_formulas = 0
        total_h_expected = 8  # rows 6 through 13

        # Check H2:H5 are empty (as required by context)
        h_early_empty = all(
            ws.cell(row=r, column=8).value is None
            for r in range(2, 6)
        )

        # Expected: H6=AVERAGE(C2:C5), H7=AVERAGE(C3:C6), ..., H(6+k)=AVERAGE(C(2+k):C(5+k))
        for i, row in enumerate(range(6, 14)):  # rows 6..13
            start_row = 2 + i   # C2, C3, ..., C9
            end_row = 5 + i     # C5, C6, ..., C12
            cell_val = ws.cell(row=row, column=8).value  # column H
            if cell_val is None:
                print(f"FAIL: Component 3 — H{row} is empty, expected AVERAGE(C{start_row}:C{end_row})")
                continue
            cell_norm = normalize_formula(str(cell_val))
            expected = f'=AVERAGE(C{start_row}:C{end_row})'
            expected_norm = normalize_formula(expected)
            if cell_norm == expected_norm:
                correct_h_formulas += 1
            else:
                print(f"FAIL: Component 3 — H{row}: expected {expected!r}, found {cell_val!r}")

        if correct_h_formulas == total_h_expected and h_early_empty:
            print(f"PASS: Component 3 — All {total_h_expected} 4-week avg formulas correct in H6:H13 (0.20 pts)")
            total_score += 0.20
        elif correct_h_formulas == total_h_expected and not h_early_empty:
            print(f"PARTIAL: Component 3 — {total_h_expected} avg formulas correct but H2:H5 not empty (0.10 pts)")
            total_score += 0.10
        elif correct_h_formulas >= 4:
            partial = round(0.20 * correct_h_formulas / total_h_expected, 4)
            print(f"PARTIAL: Component 3 — {correct_h_formulas}/{total_h_expected} avg formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {correct_h_formulas}/{total_h_expected} avg formulas correct (0 pts)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: Line chart with 2 series and title 'Weight Progress' (0.20 points)
    # Chart must: exist, be a line chart, have 2 series (actual weight + avg),
    # and have the title 'Weight Progress'
    # This FAILS on initial (no charts) and PASSES on golden
    # -------------------------------------------------------------------------
    try:
        from openpyxl.chart import LineChart

        charts = ws._charts
        if not charts:
            print("FAIL: Component 4 — No charts found on FitnessTracker sheet (0 pts)")
        else:
            found_chart = None
            for chart in charts:
                # Check it's a line chart (LineChart instance or has lineChart type)
                is_line = isinstance(chart, LineChart)
                if is_line and len(chart.series) >= 2:
                    found_chart = chart
                    break
                elif is_line:
                    found_chart = chart  # may still be valid, keep for title check

            if found_chart is None:
                # Try any chart with 2 series
                for chart in charts:
                    if len(chart.series) >= 2:
                        found_chart = chart
                        break
                if found_chart is None:
                    found_chart = charts[0]

            # Check title
            title_ok = False
            title_text = None
            try:
                title_text = found_chart.title.tx.rich.p[0].r[0].t
                if title_text and 'Weight Progress'.lower() in title_text.lower():
                    title_ok = True
            except Exception:
                pass

            has_2_series = len(found_chart.series) >= 2

            if title_ok and has_2_series and isinstance(found_chart, LineChart):
                print(f"PASS: Component 4 — Line chart with title '{title_text}' and {len(found_chart.series)} series (0.20 pts)")
                total_score += 0.20
            elif has_2_series and isinstance(found_chart, LineChart):
                print(f"PARTIAL: Component 4 — Line chart has {len(found_chart.series)} series but title check failed (title={title_text!r}) (0.10 pts)")
                total_score += 0.10
            elif title_ok and isinstance(found_chart, LineChart):
                print(f"PARTIAL: Component 4 — Line chart title correct but only {len(found_chart.series)} series (0.10 pts)")
                total_score += 0.10
            elif isinstance(found_chart, LineChart):
                print(f"PARTIAL: Component 4 — Line chart exists but title wrong and series={len(found_chart.series)} (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 4 — Chart found but not a line chart or wrong series count (0 pts)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -------------------------------------------------------------------------
    # Component 5: Conditional formatting on F3:F13 (0.10 points)
    # Green fill for negative (weight loss), Red fill for positive (weight gain)
    # This FAILS on initial (no conditional formatting) and PASSES on golden
    # -------------------------------------------------------------------------
    try:
        cf_rules = ws.conditional_formatting
        cf_found = False
        has_green_for_negative = False
        has_red_for_positive = False

        for cf_range, rules in cf_rules._cf_rules.items():
            # Look for rules that apply to F column range
            cf_str = str(cf_range)
            if 'F3' in cf_str or 'F' in cf_str:
                for rule in rules:
                    rule_type = getattr(rule, 'type', '')
                    operator = getattr(rule, 'operator', '')
                    formula = getattr(rule, 'formula', [])
                    dxf = rule.dxf
                    if dxf and dxf.fill:
                        fg_rgb = None
                        try:
                            fg_rgb = dxf.fill.fgColor.rgb
                        except Exception:
                            pass

                        # Green fill (FF00FF00) for lessThan 0 (negative = weight loss)
                        if operator == 'lessThan' and fg_rgb and 'FF00' in fg_rgb.upper() and fg_rgb.upper().startswith('FF00FF00'):
                            has_green_for_negative = True
                        elif operator == 'lessThan' and fg_rgb and fg_rgb.upper() == 'FF00FF00':
                            has_green_for_negative = True

                        # Red fill (FFFF0000) for greaterThan 0 (positive = weight gain)
                        if operator == 'greaterThan' and fg_rgb and fg_rgb.upper() == 'FFFF0000':
                            has_red_for_positive = True
                    cf_found = True

        if has_green_for_negative and has_red_for_positive:
            print("PASS: Component 5 — Conditional formatting on F3:F13: green for negative, red for positive (0.10 pts)")
            total_score += 0.10
        elif cf_found and (has_green_for_negative or has_red_for_positive):
            print(f"PARTIAL: Component 5 — Partial CF: green_neg={has_green_for_negative}, red_pos={has_red_for_positive} (0.05 pts)")
            total_score += 0.05
        elif cf_found:
            print("PARTIAL: Component 5 — Conditional formatting exists on F-range but rules don't match expected colors (0.03 pts)")
            total_score += 0.03
        else:
            print("FAIL: Component 5 — No conditional formatting found on F3:F13 (0 pts)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {round(final_score, 4)}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
