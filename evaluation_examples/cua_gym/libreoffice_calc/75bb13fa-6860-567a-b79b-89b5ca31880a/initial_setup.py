"""
Initial Setup: Fitness progress tracker over 12 weeks
Task ID: calc_gen_personal_028
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_personal_028'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: FitnessTracker ---
    ws = wb.active
    ws.title = 'FitnessTracker'

    # Headers (row 1)
    headers = ['Week', 'Date', 'Weight (lbs)', 'Body Fat %', 'Workouts', 'Weight Change', 'Trend', '4-Week Avg']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # 12 weeks of realistic fitness data (columns A-E only; F, G, H left empty)
    data = [
        # Week, Date,       Weight(lbs), Body Fat%, Workouts
        [1,  '2025-01-06',  192.4,  23.1,  3],
        [2,  '2025-01-13',  191.8,  22.9,  4],
        [3,  '2025-01-20',  190.5,  22.5,  3],
        [4,  '2025-01-27',  189.3,  22.2,  5],
        [5,  '2025-02-03',  188.7,  21.8,  4],
        [6,  '2025-02-10',  188.1,  21.5,  3],
        [7,  '2025-02-17',  187.6,  21.1,  5],
        [8,  '2025-02-24',  186.9,  20.8,  4],
        [9,  '2025-03-03',  187.4,  20.9,  3],
        [10, '2025-03-10',  186.2,  20.5,  5],
        [11, '2025-03-17',  185.5,  20.1,  4],
        [12, '2025-03-24',  184.8,  19.8,  5],
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])   # Week (int)
        ws.cell(row=r, column=2, value=row_data[1])   # Date (string)
        ws.cell(row=r, column=3, value=row_data[2])   # Weight (lbs)
        ws.cell(row=r, column=4, value=row_data[3])   # Body Fat %
        ws.cell(row=r, column=5, value=row_data[4])   # Workouts
        # Columns F (6), G (7), H (8) intentionally left empty

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
