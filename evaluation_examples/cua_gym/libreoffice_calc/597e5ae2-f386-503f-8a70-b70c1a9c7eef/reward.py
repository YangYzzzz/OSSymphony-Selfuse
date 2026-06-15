"""
Reward Script: Create named ranges for regional sales columns and formula in G2
Task ID: calc_cop_named_range_006
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Named ranges NorthSales, SouthSales, WestSales all exist  — 0.5 points
  Component 2: Named ranges reference the correct cell ranges              — 0.2 points
  Component 3: G2 contains a valid formula summing D2+E2+F2               — 0.3 points
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_cop_named_range_006'

# Expected named range definitions
EXPECTED_NAMED_RANGES = {
    'NorthSales': 'RegionalData!$D$2:$D$13',
    'SouthSales': 'RegionalData!$E$2:$E$13',
    'WestSales':  'RegionalData!$F$2:$F$13',
}


def normalize_range_ref(ref):
    """Normalize range reference for comparison: uppercase, no extra spaces."""
    return ref.strip().upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet 'RegionalData' must exist
    if 'RegionalData' not in wb.sheetnames:
        print("CRITICAL: Sheet 'RegionalData' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['RegionalData']

    # -------------------------------------------------------------------------
    # Component 1: Named ranges NorthSales, SouthSales, WestSales all exist
    #              Each named range earns 1/6 point; all three = 0.5 points
    #              Partial: 1 range = ~0.17, 2 ranges = ~0.33, 3 ranges = 0.5
    # -------------------------------------------------------------------------
    try:
        defined_names = wb.defined_names
        found_names = {}
        for range_name in ['NorthSales', 'SouthSales', 'WestSales']:
            if range_name in defined_names:
                found_names[range_name] = defined_names[range_name].attr_text
                print(f"PASS: Named range '{range_name}' exists (ref: {defined_names[range_name].attr_text})")
            else:
                print(f"FAIL: Named range '{range_name}' does not exist")

        if len(found_names) == 3:
            print(f"PASS: Component 1 — all 3 named ranges exist (0.5 pts)")
            total_score += 0.5
        elif len(found_names) == 2:
            print(f"PARTIAL: Component 1 — 2 of 3 named ranges exist (~0.33 pts)")
            total_score += 0.33
        elif len(found_names) == 1:
            print(f"PARTIAL: Component 1 — 1 of 3 named ranges exist (~0.17 pts)")
            total_score += 0.17
        else:
            print(f"FAIL: Component 1 — no named ranges found (0 pts)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        found_names = {}

    # -------------------------------------------------------------------------
    # Component 2: Named ranges reference the correct cell ranges (0.2 points)
    #              All three must map to correct column/row ranges
    # -------------------------------------------------------------------------
    try:
        defined_names = wb.defined_names
        correct_count = 0
        for range_name, expected_ref in EXPECTED_NAMED_RANGES.items():
            if range_name in defined_names:
                actual_ref = defined_names[range_name].attr_text
                if normalize_range_ref(actual_ref) == normalize_range_ref(expected_ref):
                    correct_count += 1
                    print(f"PASS: '{range_name}' references correct range {actual_ref}")
                else:
                    print(f"FAIL: '{range_name}' expected ref '{expected_ref}', got '{actual_ref}'")
            else:
                print(f"FAIL: '{range_name}' not found — cannot check reference")

        if correct_count == 3:
            print(f"PASS: Component 2 — all named ranges reference correct cells (0.2 pts)")
            total_score += 0.2
        elif correct_count == 2:
            print(f"PARTIAL: Component 2 — 2 of 3 named ranges reference correct cells (~0.13 pts)")
            total_score += 0.13
        elif correct_count == 1:
            print(f"PARTIAL: Component 2 — 1 of 3 named ranges reference correct cells (~0.07 pts)")
            total_score += 0.07
        else:
            print(f"FAIL: Component 2 — no named ranges reference correct cells (0 pts)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: G2 contains a formula summing D2+E2+F2 (0.3 points)
    #   - 0.2 pts: G2 contains any formula (not empty/raw value)
    #   - 0.3 pts: formula references D2, E2, and F2 (all three regions)
    # -------------------------------------------------------------------------
    try:
        g2_val = ws['G2'].value

        if g2_val is None:
            print(f"FAIL: Component 3 — G2 is empty (expected a formula summing D2+E2+F2)")
        elif not isinstance(g2_val, str) or not g2_val.startswith('='):
            print(f"FAIL: Component 3 — G2 is not a formula: {repr(g2_val)}")
        else:
            formula_upper = g2_val.upper().replace(' ', '')
            # Check it references D2, E2, F2 (the three regional cells)
            has_d2 = 'D2' in formula_upper
            has_e2 = 'E2' in formula_upper
            has_f2 = 'F2' in formula_upper

            if has_d2 and has_e2 and has_f2:
                print(f"PASS: Component 3 — G2 formula references D2, E2, F2: {repr(g2_val)} (0.3 pts)")
                total_score += 0.3
            elif has_d2 or has_e2 or has_f2:
                refs_found = [r for r, found in [('D2', has_d2), ('E2', has_e2), ('F2', has_f2)] if found]
                print(f"PARTIAL: Component 3 — G2 formula references {refs_found} but not all 3 cells: {repr(g2_val)} (0.2 pts)")
                total_score += 0.2
            else:
                print(f"PARTIAL: Component 3 — G2 contains a formula but doesn't reference D2/E2/F2: {repr(g2_val)} (0.1 pts)")
                total_score += 0.1
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
