"""
Initial Setup: Create RegionalData spreadsheet with monthly sales by region
Task ID: calc_cop_named_range_006
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_cop_named_range_006'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: RegionalData ---
    ws = wb.active
    ws.title = 'RegionalData'

    # Headers: Month, Budget, Actual, North, South, West, Total
    headers = ['Month', 'Budget', 'Actual', 'North', 'South', 'West', 'Total']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Monthly data rows 2-13 (Jan-Dec)
    # D2=45000, E2=38000, F2=52000 are specified; G column stays empty
    months = [
        'January', 'February', 'March', 'April', 'May', 'June',
        'July', 'August', 'September', 'October', 'November', 'December'
    ]
    # Budget, Actual, North, South, West values
    data = [
        # Month,  Budget,  Actual,  North,  South,  West
        ('January',   130000, 135000, 45000, 38000, 52000),
        ('February',  125000, 128400, 42500, 36800, 49100),
        ('March',     140000, 142300, 48200, 41500, 52600),
        ('April',     135000, 131500, 44800, 37200, 49500),
        ('May',       150000, 155800, 52000, 44100, 59700),
        ('June',      145000, 148200, 49500, 42000, 56700),
        ('July',      155000, 162400, 54300, 46200, 61900),
        ('August',    148000, 151700, 51200, 43800, 56700),
        ('September', 160000, 158900, 53600, 45700, 59600),
        ('October',   165000, 169200, 56800, 48500, 63900),
        ('November',  170000, 173600, 58100, 49700, 65800),
        ('December',  180000, 185400, 62500, 53200, 69700),
    ]

    for r, (month, budget, actual, north, south, west) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=month)
        ws.cell(row=r, column=2, value=budget)
        ws.cell(row=r, column=3, value=actual)
        ws.cell(row=r, column=4, value=north)
        ws.cell(row=r, column=5, value=south)
        ws.cell(row=r, column=6, value=west)
        # Column G (Total) is intentionally left empty — task requires formula to be added

    # Column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12

    # No named ranges — task requires agent to create them
    # No formula in G2 — task requires agent to write it

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheet: RegionalData')
    print('Rows: 1 header + 12 data rows (Jan-Dec)')
    print('Columns: Month, Budget, Actual, North, South, West, Total (G empty)')
    print('D2=45000, E2=38000, F2=52000')
    print('No named ranges defined')
    print('G column: all empty')


create_initial()
