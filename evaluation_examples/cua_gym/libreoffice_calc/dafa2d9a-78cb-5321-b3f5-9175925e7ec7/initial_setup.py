"""
Initial Setup: Build a project risk register with severity matrix formatting and status tracking.
Task ID: calc_gpm_046
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_046'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'RiskReg'

    # --- Title Row (A1:I1 merged) ---
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = 'Project Risk Register - Q2 2026'
    title_cell.font = Font(size=14, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='FF8B0000', end_color='FF8B0000', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Headers in Row 3 ---
    headers = ['Risk ID', 'Description', 'Category', 'Likelihood (1-5)',
               'Impact (1-5)', 'Risk Score', 'Priority', 'Mitigation', 'Owner']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='000000')
    all_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = all_border

    # --- Risk Data (Rows 4-13) ---
    risks = [
        ['RSK-001', 'Critical path schedule delay due to vendor integration testing overruns',
         'Schedule', 4, 5, None, None,
         'Add 2-week buffer to integration milestones; establish daily standup with vendor team',
         'Jennifer Walsh'],
        ['RSK-002', 'Budget overrun from unplanned infrastructure scaling requirements',
         'Budget', 3, 4, None, None,
         'Pre-approve contingency fund of 15%; monthly cost reviews with finance',
         'David Park'],
        ['RSK-003', 'Key senior developer resource shortage during peak delivery phase',
         'Resource', 4, 4, None, None,
         'Cross-train two junior developers; maintain contractor shortlist',
         'Maria Santos'],
        ['RSK-004', 'Scope creep from evolving stakeholder requirements mid-sprint',
         'Schedule', 3, 3, None, None,
         'Strict change control board review; bi-weekly scope alignment sessions',
         'Robert Kim'],
        ['RSK-005', 'Third-party cloud vendor service outage during production deployment',
         'External', 2, 5, None, None,
         'Multi-region failover architecture; vendor SLA enforcement with penalties',
         'Aisha Patel'],
        ['RSK-006', 'Security breach through newly discovered API vulnerability',
         'Technical', 3, 5, None, None,
         'Weekly vulnerability scans; automated dependency updates; bug bounty program',
         'Thomas Chen'],
        ['RSK-007', 'Regulatory compliance issue with upcoming GDPR data residency rules',
         'External', 2, 4, None, None,
         'Engage legal counsel early; implement data classification framework',
         'Laura Bennett'],
        ['RSK-008', 'Communication gap between distributed engineering teams across time zones',
         'Resource', 3, 3, None, None,
         'Overlapping core hours policy; async documentation standards; weekly all-hands',
         'Priya Sharma'],
        ['RSK-009', 'Legacy system technology failure during data migration phase',
         'Technical', 4, 4, None, None,
         'Incremental migration with rollback points; parallel run validation period',
         'Carlos Rivera'],
        ['RSK-010', 'Market change reducing product demand before launch date',
         'External', 2, 3, None, None,
         'Continuous market research; modular feature design for pivot capability',
         'Sarah Mitchell'],
    ]

    data_alignment = Alignment(vertical='center', wrap_text=True)

    for r, risk in enumerate(risks, 4):
        for c, val in enumerate(risk, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = all_border
            cell.alignment = data_alignment

    # --- Data Validation for D and E columns (whole number 1-5) ---
    dv_likelihood = DataValidation(
        type='whole',
        operator='between',
        formula1='1',
        formula2='5',
        allow_blank=False,
    )
    dv_likelihood.error = 'Please enter a number between 1 and 5'
    dv_likelihood.errorTitle = 'Invalid Value'
    dv_likelihood.add('D4:D13')
    ws.add_data_validation(dv_likelihood)

    dv_impact = DataValidation(
        type='whole',
        operator='between',
        formula1='1',
        formula2='5',
        allow_blank=False,
    )
    dv_impact.error = 'Please enter a number between 1 and 5'
    dv_impact.errorTitle = 'Invalid Value'
    dv_impact.add('E4:E13')
    ws.add_data_validation(dv_impact)

    # --- Column widths ---
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 16

    # --- Row height for title ---
    ws.row_dimensions[1].height = 30

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
