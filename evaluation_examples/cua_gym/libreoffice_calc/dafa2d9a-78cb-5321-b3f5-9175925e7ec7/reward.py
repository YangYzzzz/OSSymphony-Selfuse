"""
Reward Script: Build a project risk register with severity matrix formatting and status tracking.
Task ID: calc_gpm_046
Domain: libreoffice_calc
Scoring:
  Component 1: Risk Score formulas in F4:F13 (0.25)
  Component 2: Priority IF formulas in G4:G13 (0.25)
  Component 3: Risk Summary section rows 15-18 (0.20)
  Component 4: Category dropdown data validation on C4:C13 (0.15)
  Component 5: Conditional formatting on F and G columns (0.15)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_046'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print("CRITICAL: Cannot load file {}: {}".format(file_path, e))
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'RiskReg' sheet must exist
    if 'RiskReg' not in wb.sheetnames:
        print("FAIL: Sheet 'RiskReg' not found. Sheets: {}".format(wb.sheetnames))
        print("REWARD: 0.0")
        return 0.0

    ws = wb['RiskReg']

    # Component 1: Risk Score formulas =D*E in F4:F13 (0.25 points)
    try:
        formula_count = 0
        for row in range(4, 14):
            val = ws.cell(row=row, column=6).value  # F column
            if val is not None and isinstance(val, str):
                normalized = val.upper().replace(" ", "")
                expected = "=D{}*E{}".format(row, row).upper()
                if normalized == expected:
                    formula_count += 1
        if formula_count == 10:
            print("PASS: Component 1 - All 10 Risk Score formulas (=D*E) found in F4:F13 (0.25 pts)")
            total_score += 0.25
        elif formula_count >= 5:
            partial = 0.25 * (formula_count / 10.0)
            print("PARTIAL: Component 1 - {}/10 Risk Score formulas found ({}pts)".format(formula_count, round(partial, 3)))
            total_score += partial
        else:
            print("FAIL: Component 1 - Only {}/10 Risk Score formulas found in F4:F13".format(formula_count))
    except Exception as e:
        print("ERROR: Component 1 - {}".format(e))

    # Component 2: Priority IF formulas in G4:G13 (0.25 points)
    try:
        if_count = 0
        for row in range(4, 14):
            val = ws.cell(row=row, column=7).value  # G column
            if val is not None and isinstance(val, str):
                normalized = val.upper().replace(" ", "")
                # Check that it's an IF formula referencing F column with Critical/High/Medium/Low
                if normalized.startswith("=IF(") and "CRITICAL" in normalized and "HIGH" in normalized and "MEDIUM" in normalized and "LOW" in normalized:
                    if_count += 1
        if if_count == 10:
            print("PASS: Component 2 - All 10 Priority IF formulas found in G4:G13 (0.25 pts)")
            total_score += 0.25
        elif if_count >= 5:
            partial = 0.25 * (if_count / 10.0)
            print("PARTIAL: Component 2 - {}/10 Priority IF formulas found ({}pts)".format(if_count, round(partial, 3)))
            total_score += partial
        else:
            print("FAIL: Component 2 - Only {}/10 Priority IF formulas found in G4:G13".format(if_count))
    except Exception as e:
        print("ERROR: Component 2 - {}".format(e))

    # Component 3: Risk Summary section rows 15-18 (0.20 points)
    # Sub-checks: A15 merged with 'Risk Summary', F15-F18 labels, G15-G18 COUNTIF formulas
    try:
        comp3_score = 0.0

        # Sub-check 3a: A15:E15 merged with 'Risk Summary' (0.08 pts)
        a15_merged = any('A15' in str(m) for m in ws.merged_cells.ranges)
        a15_val = ws.cell(row=15, column=1).value
        if a15_merged and a15_val is not None and 'Risk Summary' in str(a15_val):
            print("  PASS: 3a - A15:E15 merged with 'Risk Summary' text")
            comp3_score += 0.08
        else:
            print("  FAIL: 3a - A15 merged={}, value={}".format(a15_merged, repr(a15_val)))

        # Sub-check 3b: F15-F18 labels (Critical, High, Medium, Low) (0.04 pts)
        expected_labels = ['Critical:', 'High:', 'Medium:', 'Low:']
        label_count = 0
        for i, row in enumerate(range(15, 19)):
            val = ws.cell(row=row, column=6).value
            if val is not None and expected_labels[i].lower().replace(':', '') in str(val).lower():
                label_count += 1
        if label_count >= 4:
            print("  PASS: 3b - All 4 risk category labels found in F15:F18")
            comp3_score += 0.04
        else:
            print("  FAIL: 3b - Only {}/4 labels found in F15:F18".format(label_count))

        # Sub-check 3c: G15-G18 COUNTIF formulas (0.08 pts)
        countif_count = 0
        for row in range(15, 19):
            val = ws.cell(row=row, column=7).value
            if val is not None and isinstance(val, str) and 'COUNTIF' in val.upper():
                countif_count += 1
        if countif_count >= 4:
            print("  PASS: 3c - All 4 COUNTIF formulas found in G15:G18")
            comp3_score += 0.08
        else:
            print("  FAIL: 3c - Only {}/4 COUNTIF formulas found in G15:G18".format(countif_count))

        if comp3_score > 0:
            print("PASS: Component 3 - Risk Summary section ({}pts)".format(round(comp3_score, 3)))
            total_score += comp3_score
        else:
            print("FAIL: Component 3 - Risk Summary section not found")
    except Exception as e:
        print("ERROR: Component 3 - {}".format(e))

    # Component 4: Category dropdown data validation on C4:C13 (0.15 points)
    try:
        cat_dv_found = False
        for dv in ws.data_validations.dataValidation:
            if dv.type == 'list':
                # Check if it covers C column range
                sqref_str = str(dv.sqref)
                if 'C' in sqref_str:
                    # Check formula contains the expected categories
                    formula = str(dv.formula1).upper() if dv.formula1 else ''
                    has_categories = all(cat in formula for cat in ['TECHNICAL', 'SCHEDULE', 'BUDGET', 'RESOURCE', 'EXTERNAL'])
                    if has_categories:
                        cat_dv_found = True
                        break
        if cat_dv_found:
            print("PASS: Component 4 - Category dropdown validation on C column with correct categories (0.15 pts)")
            total_score += 0.15
        else:
            print("FAIL: Component 4 - Category dropdown validation not found on C column")
    except Exception as e:
        print("ERROR: Component 4 - {}".format(e))

    # Component 5: Conditional formatting on F4:F13 and G4:G13 (0.15 points)
    try:
        comp5_score = 0.0

        f_cf_found = False
        g_cf_found = False

        for cf in ws.conditional_formatting:
            range_str = str(cf)
            for rule in cf.rules:
                # Check for color scale on F column
                if 'F' in range_str and rule.type == 'colorScale':
                    f_cf_found = True
                # Check for expression rules on G column
                if 'G' in range_str and rule.type == 'expression':
                    g_cf_found = True

        if f_cf_found:
            print("  PASS: 5a - Color scale conditional formatting found on F column")
            comp5_score += 0.075
        else:
            print("  FAIL: 5a - No color scale conditional formatting on F column")

        if g_cf_found:
            print("  PASS: 5b - Expression-based conditional formatting found on G column")
            comp5_score += 0.075
        else:
            print("  FAIL: 5b - No expression-based conditional formatting on G column")

        if comp5_score > 0:
            print("PASS: Component 5 - Conditional formatting ({}pts)".format(round(comp5_score, 3)))
            total_score += comp5_score
        else:
            print("FAIL: Component 5 - No conditional formatting found")
    except Exception as e:
        print("ERROR: Component 5 - {}".format(e))

    final_score = round(min(total_score, 1.0), 2)
    print("\nScore: {}/1.0".format(round(total_score, 4)))
    print("REWARD: {}".format(final_score))
    return final_score


# Entry point
file_path = '{}/{}.xlsx'.format(WORKDIR, TASK_ID)
if not os.path.exists(file_path):
    print("File not found: {}".format(file_path))
    print("REWARD: 0.0")
else:
    verify_task(file_path)
