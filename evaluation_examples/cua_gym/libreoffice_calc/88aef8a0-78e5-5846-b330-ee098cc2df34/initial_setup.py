"""
Initial Setup: Acceptance Sampling Decision Tool — pre-task state
Task ID: calc_ops_qc_acceptance_sampling_024
Domain: libreoffice_calc

Creates a spreadsheet with 30 incoming shipment lots. Columns A-E are filled with
realistic data; columns F (Acceptance Number) and G (Decision) are intentionally empty
so the agent must add the formulas.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_qc_acceptance_sampling_024'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'AcceptanceSampling'

    # --- Headers (Row 1) ---
    headers = ['Lot ID', 'Supplier', 'Lot Size', 'Sample Size', 'Defectives Found',
               'Acceptance Number', 'Decision']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, name='Calibri', size=11, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Data rows (Rows 2-31): 30 lots ---
    # Realistic supplier names and lot data
    suppliers = [
        'AlphaComponents Ltd', 'BetaTech Industries', 'Gamma Manufacturing',
        'Delta Precision Parts', 'Epsilon Electronics', 'Zeta Supply Co',
        'Eta Industrial', 'Theta Components', 'Iota Materials', 'Kappa Global'
    ]

    # Each tuple: (LotID, Supplier, LotSize, SampleSize, DefectivesFound)
    # LotSize ranges 500-5000, SampleSize ~5-10% of LotSize, Defectives 0-15
    lot_data = [
        ('LOT-2025-001', suppliers[0], 2000, 80, 3),
        ('LOT-2025-002', suppliers[1], 1500, 60, 1),
        ('LOT-2025-003', suppliers[2], 3000, 120, 5),
        ('LOT-2025-004', suppliers[3], 800,  40, 0),
        ('LOT-2025-005', suppliers[4], 5000, 200, 8),
        ('LOT-2025-006', suppliers[5], 1200, 50, 2),
        ('LOT-2025-007', suppliers[6], 2500, 100, 4),
        ('LOT-2025-008', suppliers[7], 600,  30, 1),
        ('LOT-2025-009', suppliers[8], 4000, 160, 7),
        ('LOT-2025-010', suppliers[9], 900,  40, 0),
        ('LOT-2025-011', suppliers[0], 3500, 140, 6),
        ('LOT-2025-012', suppliers[1], 700,  30, 2),
        ('LOT-2025-013', suppliers[2], 2200, 90, 3),
        ('LOT-2025-014', suppliers[3], 1800, 72, 1),
        ('LOT-2025-015', suppliers[4], 4500, 180, 9),
        ('LOT-2025-016', suppliers[5], 1000, 50, 0),
        ('LOT-2025-017', suppliers[6], 3200, 128, 5),
        ('LOT-2025-018', suppliers[7], 500,  25, 1),
        ('LOT-2025-019', suppliers[8], 2800, 112, 4),
        ('LOT-2025-020', suppliers[9], 1600, 64, 2),
        ('LOT-2025-021', suppliers[0], 4200, 168, 10),
        ('LOT-2025-022', suppliers[1], 750,  30, 0),
        ('LOT-2025-023', suppliers[2], 2600, 104, 3),
        ('LOT-2025-024', suppliers[3], 1300, 52, 1),
        ('LOT-2025-025', suppliers[4], 3800, 152, 8),
        ('LOT-2025-026', suppliers[5], 550,  22, 0),
        ('LOT-2025-027', suppliers[6], 2100, 84, 2),
        ('LOT-2025-028', suppliers[7], 4800, 192, 11),
        ('LOT-2025-029', suppliers[8], 1100, 44, 1),
        ('LOT-2025-030', suppliers[9], 3300, 132, 6),
    ]

    thin_side = Side(style='thin', color='BFBFBF')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for r, (lot_id, supplier, lot_size, sample_size, defectives) in enumerate(lot_data, 2):
        ws.cell(row=r, column=1, value=lot_id).alignment = Alignment(horizontal='left')
        ws.cell(row=r, column=2, value=supplier).alignment = Alignment(horizontal='left')
        ws.cell(row=r, column=3, value=lot_size).alignment = Alignment(horizontal='right')
        ws.cell(row=r, column=4, value=sample_size).alignment = Alignment(horizontal='right')
        ws.cell(row=r, column=5, value=defectives).alignment = Alignment(horizontal='right')
        # Columns F and G are intentionally left empty (agent must fill them)

        # Apply thin border to data cells A-E
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = thin_border

    # --- Column widths ---
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 14

    # --- Freeze header row ---
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: AcceptanceSampling')
    print(f'  Rows: 1 header + 30 data rows')
    print(f'  Columns A-E filled; F (Acceptance Number) and G (Decision) intentionally empty')


create_initial()
