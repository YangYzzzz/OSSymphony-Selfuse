"""
Reward Script: Create an acceptance sampling decision tool
Task ID: calc_ops_qc_acceptance_sampling_024
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Column F (F2:F31) has ROUND(Dn*0.025,0) acceptance number formulas     — 0.30 pts
  Component 2: Column G (G2:G31) has IF(En<=Fn,"Accept","Reject") decision formulas   — 0.30 pts
  Component 3: Summary section in rows 33-36 (label + COUNTIF + rate formula)         — 0.20 pts
  Component 4: Conditional formatting on G2:G31 (green Accept / red Reject)           — 0.10 pts
  Component 5: Comment on cell F1 explaining the AQL formula                          — 0.10 pts
  Total: 1.00
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_qc_acceptance_sampling_024'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet exists as a precondition gate
    if 'AcceptanceSampling' not in wb.sheetnames:
        print("CRITICAL: Sheet 'AcceptanceSampling' not found.")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['AcceptanceSampling']

    # Component 1: Column F (F2:F31) has ROUND acceptance number formulas (0.30 points)
    # These formulas MUST NOT exist in the initial file — they are the task-introduced change.
    try:
        f_formula_count = 0
        f_formula_correct = 0
        for row in range(2, 32):
            f_val = ws.cell(row=row, column=6).value
            if f_val is not None:
                f_formula_count += 1
                f_str = str(f_val).strip().upper().replace(' ', '')
                # Expect =ROUND(Dn*0.025,0) where n = row number
                expected = f'=ROUND(D{row}*0.025,0)'.upper()
                if f_str == expected:
                    f_formula_correct += 1
        if f_formula_correct == 30:
            print(f"PASS: Component 1 — All 30 F2:F31 cells have correct ROUND(Dn*0.025,0) formulas (0.30 pts)")
            total_score += 0.30
        elif f_formula_correct >= 20:
            print(f"PARTIAL: Component 1 — {f_formula_correct}/30 F column formulas correct (0.15 pts)")
            total_score += 0.15
        elif f_formula_correct >= 10:
            print(f"PARTIAL: Component 1 — {f_formula_correct}/30 F column formulas correct (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 — Only {f_formula_correct}/30 F column formulas correct (found {f_formula_count} non-empty). Expected =ROUND(Dn*0.025,0)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Column G (G2:G31) has IF decision formulas (0.30 points)
    # These formulas MUST NOT exist in the initial file — they are task-introduced changes.
    try:
        g_formula_count = 0
        g_formula_correct = 0
        for row in range(2, 32):
            g_val = ws.cell(row=row, column=7).value
            if g_val is not None:
                g_formula_count += 1
                g_str = str(g_val).strip().upper().replace(' ', '').replace('"', "'")
                # Expect =IF(En<=Fn,"Accept","Reject") — check with both quote styles
                g_str_dq = str(g_val).strip().upper().replace(' ', '')
                expected_dq = f'=IF(E{row}<=F{row},"ACCEPT","REJECT")'
                expected_sq = f"=IF(E{row}<=F{row},'ACCEPT','REJECT')"
                if g_str_dq == expected_dq or g_str.replace('"', "'") == expected_sq.upper():
                    g_formula_correct += 1
        if g_formula_correct == 30:
            print(f"PASS: Component 2 — All 30 G2:G31 cells have correct IF(En<=Fn,Accept,Reject) formulas (0.30 pts)")
            total_score += 0.30
        elif g_formula_correct >= 20:
            print(f"PARTIAL: Component 2 — {g_formula_correct}/30 G column formulas correct (0.15 pts)")
            total_score += 0.15
        elif g_formula_correct >= 10:
            print(f"PARTIAL: Component 2 — {g_formula_correct}/30 G column formulas correct (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 — Only {g_formula_correct}/30 G column formulas correct. Expected =IF(En<=Fn,\"Accept\",\"Reject\")")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Summary section below row 32 (0.20 points)
    # Must contain: Summary label, COUNTIF for Accept/Reject lots, acceptance rate formula
    try:
        summary_score = 0.0

        # Find summary label (should be in row 33 col A, but allow slight variation rows 33-40)
        summary_row = None
        for r in range(32, 42):
            cell_val = ws.cell(row=r, column=1).value
            if cell_val and str(cell_val).strip().lower() == 'summary':
                summary_row = r
                break

        if summary_row is not None:
            summary_score += 0.05
            # Look for COUNTIF("Accept") in the rows below summary
            accepted_row = None
            rejected_row = None
            rate_row = None
            for r in range(summary_row + 1, summary_row + 6):
                label_val = ws.cell(row=r, column=1).value
                b_val = ws.cell(row=r, column=2).value
                if label_val:
                    label_lower = str(label_val).lower()
                    if 'accept' in label_lower and b_val and 'COUNTIF' in str(b_val).upper() and 'ACCEPT' in str(b_val).upper():
                        accepted_row = r
                        summary_score += 0.05
                    elif 'reject' in label_lower and b_val and 'COUNTIF' in str(b_val).upper() and 'REJECT' in str(b_val).upper():
                        rejected_row = r
                        summary_score += 0.05
                    elif ('rate' in label_lower or 'acceptance' in label_lower) and b_val is not None:
                        rate_row = r
                        # Check it's a formula or numeric value
                        b_str = str(b_val).strip()
                        if b_str.startswith('=') or b_str.replace('.', '').replace('/', '').isdigit():
                            summary_score += 0.05

            if summary_score >= 0.20:
                print(f"PASS: Component 3 — Summary section found with label, COUNTIF(Accept), COUNTIF(Reject), and rate formula (0.20 pts)")
                total_score += 0.20
            elif summary_score >= 0.10:
                partial = round(summary_score, 2)
                print(f"PARTIAL: Component 3 — Summary section partially complete (score={partial}): summary_row={summary_row}, accepted_row={accepted_row}, rejected_row={rejected_row}, rate_row={rate_row}")
                total_score += partial
            else:
                if summary_score > 0:
                    print(f"FAIL: Component 3 — Summary section incomplete (score={summary_score}): summary_row={summary_row}")
                    total_score += summary_score
                else:
                    print(f"FAIL: Component 3 — Summary section incomplete, score=0")
        else:
            print(f"FAIL: Component 3 — No 'Summary' label found in rows 32-41 column A")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Conditional formatting on G2:G31 (green Accept / red Reject) (0.10 points)
    try:
        cf_rules = ws.conditional_formatting
        accept_green_count = 0
        reject_red_count = 0

        for cf_range in cf_rules:
            cf_range_str = str(cf_range)
            # Check if the range covers G2:G31 (or similar)
            if 'G' in cf_range_str.upper():
                for rule in cf_range.rules:
                    try:
                        formula_str = ' '.join(str(f) for f in (rule.formula or []))
                        if 'ACCEPT' in formula_str.upper():
                            # Should have green fill
                            if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                                rgb = rule.dxf.fill.fgColor.rgb
                                # Green-ish color (common greens: 92D050, 00FF00, 70AD47, etc.)
                                if rgb and (rgb.upper().startswith('FF9') or rgb.upper() in ('FF92D050', 'FF00FF00', 'FF70AD47', 'FF00B050')):
                                    accept_green_count += 1
                        if 'REJECT' in formula_str.upper():
                            # Should have red fill
                            if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                                rgb = rule.dxf.fill.fgColor.rgb
                                # Red-ish color (R >= 180, G < 100, B < 100)
                                if rgb and len(rgb) == 8:
                                    try:
                                        r_val = int(rgb[2:4], 16)
                                        g_val = int(rgb[4:6], 16)
                                        b_val = int(rgb[6:8], 16)
                                        if r_val >= 180 and g_val < 100 and b_val < 100:
                                            reject_red_count += 1
                                    except Exception:
                                        pass
                    except Exception as inner_e:
                        pass

        if accept_green_count >= 1 and reject_red_count >= 1:
            print(f"PASS: Component 4 — Conditional formatting: green for Accept and red for Reject on G column (0.10 pts)")
            total_score += 0.10
        elif accept_green_count >= 1 or reject_red_count >= 1:
            print(f"PARTIAL: Component 4 — Conditional formatting partially set: accept_green={accept_green_count}, reject_red={reject_red_count} (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4 — No conditional formatting with green/red on G column found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Comment on cell F1 explaining the AQL formula (0.10 points)
    try:
        f1_cell = ws['F1']
        if hasattr(f1_cell, 'comment') and f1_cell.comment is not None:
            comment_text = f1_cell.comment.text if f1_cell.comment.text else ''
            # Check that comment contains relevant AQL keywords
            comment_lower = comment_text.lower()
            has_aql = 'aql' in comment_lower or 'acceptable quality' in comment_lower or 'acceptance' in comment_lower
            has_formula = '0.025' in comment_text or 'round' in comment_lower or '2.5%' in comment_text or '2.5' in comment_text
            if has_aql and has_formula:
                print(f"PASS: Component 5 — F1 has comment explaining AQL formula (0.10 pts): '{comment_text[:80]}'")
                total_score += 0.10
            elif comment_text.strip():
                print(f"PARTIAL: Component 5 — F1 has comment but may be incomplete (0.05 pts): '{comment_text[:80]}'")
                total_score += 0.05
            else:
                print(f"FAIL: Component 5 — F1 has empty comment")
        else:
            print(f"FAIL: Component 5 — F1 has no comment")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {round(final_score, 4)}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
