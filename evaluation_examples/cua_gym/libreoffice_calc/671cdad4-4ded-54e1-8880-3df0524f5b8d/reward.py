"""
Reward Script: Real estate listing sheet cleanup
Task ID: calc_gen_data_cleanup_014
Domain: libreoffice_calc
Scoring:
  Component 1: No merged cells remain (0.25 pts)
  Component 2: All blank rows deleted — 73 data rows remain (0.25 pts)
  Component 3: Price column (G) all converted to numeric integers (0.30 pts)
  Component 4: Table style applied (alternating row shading) and AutoFilter enabled (0.20 pts)
Total: 1.0
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_data_cleanup_014'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Listings' sheet must exist
    if 'Listings' not in wb.sheetnames:
        print("CRITICAL: 'Listings' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Listings']

    # -----------------------------------------------------------------------
    # Component 1: No merged cells remain (0.25 points)
    # In the initial file there are 14 merged cell ranges; task requires all
    # to be unmerged so every cell holds its own value.
    # -----------------------------------------------------------------------
    try:
        merged_count = len(list(ws.merged_cells.ranges))
        if merged_count == 0:
            print(f"PASS: Component 1 — No merged cells (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — {merged_count} merged cell range(s) still present (expected 0)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: All blank rows deleted — 73 data rows remain (0.25 points)
    # Initial file has 81 rows (row 1 = header + 80 rows with 7 blank rows).
    # After deletion: 1 header row + 73 data rows = 74 total rows.
    # We check that no completely blank data rows remain and row count is ~73.
    # -----------------------------------------------------------------------
    try:
        # Count blank rows (all 8 columns None) in data range
        blank_row_count = 0
        data_row_count = 0
        for row in range(2, ws.max_row + 1):
            row_vals = [ws.cell(row=row, column=col).value for col in range(1, 9)]
            if all(v is None for v in row_vals):
                blank_row_count += 1
            else:
                data_row_count += 1

        if blank_row_count == 0 and data_row_count == 73:
            print(f"PASS: Component 2 — No blank rows, {data_row_count} data rows present (0.25 pts)")
            total_score += 0.25
        elif blank_row_count == 0 and data_row_count > 0:
            # All blanks removed but different row count — partial pass not awarded
            print(f"FAIL: Component 2 — No blank rows but {data_row_count} data rows (expected 73)")
        else:
            print(f"FAIL: Component 2 — {blank_row_count} blank row(s) still present, {data_row_count} data rows")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Price column (G) all converted to numeric integers (0.30 points)
    # Initial file has text values like '$785,000', '$1.2M' in column G.
    # After task: all should be real numeric values (int or float).
    # Also verify number_format is currency ($#,##0).
    # -----------------------------------------------------------------------
    try:
        non_numeric_g = []
        numeric_g = []
        none_g = []
        wrong_format_g = []

        for row in range(2, ws.max_row + 1):
            row_vals = [ws.cell(row=row, column=col).value for col in range(1, 9)]
            # Skip blank rows (already handled in component 2)
            if all(v is None for v in row_vals):
                continue

            cell = ws.cell(row=row, column=7)
            val = cell.value
            fmt = cell.number_format

            if val is None:
                none_g.append(row)
            elif isinstance(val, (int, float)):
                numeric_g.append(row)
                # Also check format
                if fmt != '$#,##0':
                    wrong_format_g.append((row, fmt))
            else:
                non_numeric_g.append((row, val))

        if len(non_numeric_g) == 0 and len(numeric_g) > 0:
            # All non-blank G values are numeric
            if len(wrong_format_g) == 0:
                print(f"PASS: Component 3 — All {len(numeric_g)} price values are numeric with $#,##0 format (0.30 pts)")
                total_score += 0.30
            else:
                # Numeric conversion done but format not applied — award partial
                # (No partial sub-score within a component; require both to be complete)
                print(f"FAIL: Component 3 — Values are numeric but {len(wrong_format_g)} cell(s) have wrong format "
                      f"(e.g. row {wrong_format_g[0][0]}: {repr(wrong_format_g[0][1])})")
        else:
            print(f"FAIL: Component 3 — {len(non_numeric_g)} non-numeric price value(s) remain "
                  f"(e.g. {non_numeric_g[:3]})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: Table style applied + AutoFilter enabled (0.20 points)
    # Golden file uses alternating row shading (FFFFFFFF / FFD6E4F0) and has
    # AutoFilter set on A1:H<last_row>.
    # Award 0.10 for alternating fill and 0.10 for AutoFilter.
    # -----------------------------------------------------------------------
    sub_score_4 = 0.0

    # Sub-component 4a: AutoFilter enabled (0.10 pts)
    try:
        if ws.auto_filter.ref:
            print(f"PASS: Component 4a — AutoFilter enabled: {ws.auto_filter.ref} (0.10 pts)")
            sub_score_4 += 0.10
        else:
            print("FAIL: Component 4a — AutoFilter not set on the sheet")
    except Exception as e:
        print(f"ERROR: Component 4a — {e}")

    # Sub-component 4b: Alternating row shading applied (0.10 pts)
    # Check that data rows have a fill pattern applied (not all plain/no fill)
    # The golden uses: even data rows = FFFFFFFF, odd data rows = FFD6E4F0
    try:
        shaded_rows = 0
        plain_rows = 0
        sampled = 0
        for row in range(2, min(ws.max_row + 1, 12)):
            row_vals = [ws.cell(row=row, column=col).value for col in range(1, 9)]
            if all(v is None for v in row_vals):
                continue
            cell_a = ws.cell(row=row, column=1)
            try:
                rgb = cell_a.fill.fgColor.rgb
                fill_type = cell_a.fill.fill_type
                if fill_type == 'solid' and rgb not in ('00000000', None):
                    shaded_rows += 1
                else:
                    plain_rows += 1
            except Exception:
                plain_rows += 1
            sampled += 1

        if sampled > 0 and shaded_rows >= sampled // 2:
            print(f"PASS: Component 4b — Alternating row shading applied "
                  f"({shaded_rows}/{sampled} sampled rows have fill) (0.10 pts)")
            sub_score_4 += 0.10
        else:
            print(f"FAIL: Component 4b — Table/alternating shading not detected "
                  f"({shaded_rows}/{sampled} sampled rows with fill)")
    except Exception as e:
        print(f"ERROR: Component 4b — {e}")

    if sub_score_4 > 0:
        total_score += sub_score_4

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
