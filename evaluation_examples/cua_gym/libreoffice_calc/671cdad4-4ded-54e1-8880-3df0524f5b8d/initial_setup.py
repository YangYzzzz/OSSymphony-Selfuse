"""
Initial Setup: Real estate listing sheet with messy data for cleanup
Task ID: calc_gen_data_cleanup_014
Domain: libreoffice_calc

Creates a messy real estate listings spreadsheet with:
- Merged cells in address/city area
- Blank rows scattered throughout
- Price column with text formats ($1.2M, $450,000 as text) and real numbers mixed
- No table formatting or AutoFilter
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_data_cleanup_014'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Listings'

    # --- Headers ---
    headers = ['MLS #', 'Address', 'City', 'Beds', 'Baths', 'Sqft', 'Price', 'Agent']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # --- Realistic listing data (70 rows with 10 blank rows interspersed) ---
    # Data: MLS#, Address, City, Beds, Baths, Sqft, Price (mixed formats), Agent
    listings_data = [
        ['MLS-10142', '4521 Willow Creek Dr', 'Austin', 4, 3, 2450, '$785,000', 'Jennifer Walsh'],
        ['MLS-10143', '8830 Sunset Blvd', 'Austin', 3, 2, 1820, '$542,000', 'Robert Hawkins'],
        ['MLS-10144', '112 Lakewood Ct', 'Round Rock', 5, 4, 3200, '$1.1M', 'Sarah Chen'],
        ['MLS-10145', '7741 Maple Ave', 'Cedar Park', 3, 2, 1650, '$468,500', 'Marcus Johnson'],
        ['MLS-10146', '295 River Oaks Ln', 'Georgetown', 4, 3, 2780, '$895,000', 'Lisa Tanaka'],
        # blank row 1
        ['MLS-10147', '663 Pecan Grove Rd', 'Pflugerville', 3, 2, 1720, '$399,000', 'David Kim'],
        ['MLS-10148', '1480 Highland Hills Dr', 'Austin', 6, 5, 4100, '$1.5M', 'Jennifer Walsh'],
        ['MLS-10149', '52 Mockingbird Ln', 'Leander', 4, 3, 2310, '$612,000', 'Sarah Chen'],
        ['MLS-10150', '3390 Oak Trail Blvd', 'Austin', 3, 2, 1890, '$559,000', 'Robert Hawkins'],
        ['MLS-10151', '878 Cypress Mill Rd', 'Cedar Park', 4, 3, 2620, '$739,500', 'Marcus Johnson'],
        ['MLS-10152', '45 Bluebell Cir', 'Round Rock', 3, 2, 1540, '$378,000', 'Lisa Tanaka'],
        ['MLS-10153', '2211 Lakeview Dr', 'Austin', 5, 4, 3560, '$1.2M', 'David Kim'],
        # blank row 2
        ['MLS-10154', '730 Santa Fe Trl', 'Georgetown', 4, 3, 2190, '$625,000', 'Jennifer Walsh'],
        ['MLS-10155', '1605 Rolling Hills Dr', 'Pflugerville', 3, 2, 1710, '$415,000', 'Sarah Chen'],
        ['MLS-10156', '88 Magnolia Park Ct', 'Leander', 4, 4, 2850, '$798,000', 'Robert Hawkins'],
        ['MLS-10157', '5543 Longhorn Pass', 'Austin', 3, 2, 1980, '$583,000', 'Marcus Johnson'],
        ['MLS-10158', '411 Deer Creek Rd', 'Cedar Park', 4, 3, 2430, '$710,000', 'Lisa Tanaka'],
        ['MLS-10159', '1900 Vista Ridge Pkwy', 'Leander', 5, 4, 3310, '$1.05M', 'David Kim'],
        ['MLS-10160', '67 Clearwater Springs', 'Round Rock', 3, 2, 1620, '$388,500', 'Jennifer Walsh'],
        # blank row 3
        ['MLS-10161', '340 Primrose Path', 'Austin', 4, 3, 2510, '$821,000', 'Sarah Chen'],
        ['MLS-10162', '1233 Wolf Ranch Pkwy', 'Georgetown', 3, 2, 1790, '$490,000', 'Robert Hawkins'],
        ['MLS-10163', '756 Balcones Dr', 'Austin', 5, 4, 3720, '$1.35M', 'Marcus Johnson'],
        ['MLS-10164', '29 Cardinal Cove', 'Pflugerville', 3, 2, 1680, '$405,000', 'Lisa Tanaka'],
        ['MLS-10165', '4490 Ironwood Cir', 'Cedar Park', 4, 3, 2680, '$759,000', 'David Kim'],
        ['MLS-10166', '1850 Lone Star Ranch Rd', 'Leander', 6, 5, 4200, '$1.8M', 'Jennifer Walsh'],
        ['MLS-10167', '122 Heritage Oak Dr', 'Round Rock', 3, 2, 1560, '$362,000', 'Sarah Chen'],
        # blank row 4
        ['MLS-10168', '980 Shoal Creek Blvd', 'Austin', 4, 3, 2340, '$785,000', 'Robert Hawkins'],
        ['MLS-10169', '2677 Chisholm Trail', 'Georgetown', 3, 2, 1820, '$452,000', 'Marcus Johnson'],
        ['MLS-10170', '55 Wisteria Way', 'Pflugerville', 4, 3, 2150, '$589,000', 'Lisa Tanaka'],
        ['MLS-10171', '1420 Ridgeline Dr', 'Austin', 5, 4, 3450, '$1.25M', 'David Kim'],
        ['MLS-10172', '308 Cottonwood Creek', 'Cedar Park', 3, 2, 1640, '$393,500', 'Jennifer Walsh'],
        ['MLS-10173', '6621 Travis Heights Blvd', 'Austin', 4, 3, 2670, '$921,000', 'Sarah Chen'],
        # blank row 5
        ['MLS-10174', '175 Covered Bridge Dr', 'Leander', 4, 3, 2480, '$679,000', 'Robert Hawkins'],
        ['MLS-10175', '840 Stone Canyon Rd', 'Round Rock', 3, 2, 1740, '$419,000', 'Marcus Johnson'],
        ['MLS-10176', '3322 Bull Creek Rd', 'Austin', 5, 4, 3280, '$1.15M', 'Lisa Tanaka'],
        ['MLS-10177', '70 Quail Hollow Ln', 'Georgetown', 3, 2, 1590, '$365,000', 'David Kim'],
        ['MLS-10178', '2090 Lakeline Blvd', 'Cedar Park', 4, 3, 2560, '$718,000', 'Jennifer Walsh'],
        ['MLS-10179', '1331 Barton Creek Blvd', 'Austin', 6, 5, 4650, '$2.1M', 'Sarah Chen'],
        ['MLS-10180', '449 Greenway Plaza', 'Pflugerville', 3, 2, 1680, '$412,000', 'Robert Hawkins'],
        # blank row 6
        ['MLS-10181', '5540 Mesa Vista Dr', 'Austin', 4, 3, 2390, '$765,000', 'Marcus Johnson'],
        ['MLS-10182', '228 Buttonwood Dr', 'Leander', 3, 2, 1610, '$371,000', 'Lisa Tanaka'],
        ['MLS-10183', '1760 Reunion Ranch Rd', 'Georgetown', 5, 4, 3180, '$875,000', 'David Kim'],
        ['MLS-10184', '93 Robin Hood Trl', 'Round Rock', 3, 2, 1550, '$348,000', 'Jennifer Walsh'],
        ['MLS-10185', '4710 Shoalwood Ave', 'Austin', 4, 3, 2290, '$742,000', 'Sarah Chen'],
        ['MLS-10186', '660 Fossil Creek Dr', 'Cedar Park', 4, 3, 2440, '$698,000', 'Robert Hawkins'],
        # blank row 7
        ['MLS-10187', '1120 Canyon Rim Dr', 'Austin', 5, 4, 3620, '$1.4M', 'Marcus Johnson'],
        ['MLS-10188', '37 Meadow Lark Ln', 'Pflugerville', 3, 2, 1700, '$398,000', 'Lisa Tanaka'],
        ['MLS-10189', '2980 Lime Rock Dr', 'Leander', 4, 3, 2320, '$635,000', 'David Kim'],
        ['MLS-10190', '810 Possum Hollow Rd', 'Georgetown', 3, 2, 1830, '$456,000', 'Jennifer Walsh'],
        ['MLS-10191', '5205 Far West Blvd', 'Austin', 4, 3, 2480, '$812,000', 'Sarah Chen'],
        ['MLS-10192', '163 Deer Valley Ct', 'Round Rock', 3, 2, 1560, '$355,000', 'Robert Hawkins'],
        ['MLS-10193', '3880 Slaughter Ln', 'Austin', 5, 4, 3510, '$1.3M', 'Marcus Johnson'],
        # blank row 8
        ['MLS-10194', '720 Windy Hill Rd', 'Cedar Park', 4, 3, 2250, '$665,000', 'Lisa Tanaka'],
        ['MLS-10195', '1567 Old Spicewood Rd', 'Leander', 3, 2, 1760, '$428,000', 'David Kim'],
        ['MLS-10196', '48 Peregrine Ct', 'Georgetown', 4, 3, 2600, '$729,000', 'Jennifer Walsh'],
        ['MLS-10197', '6150 Burnet Rd', 'Austin', 4, 3, 2420, 775000, 'Sarah Chen'],
        ['MLS-10198', '290 Westgate Blvd', 'Austin', 3, 2, 1690, '$435,000', 'Robert Hawkins'],
        ['MLS-10199', '1890 Parmer Ln', 'Pflugerville', 5, 4, 3340, '$945,000', 'Marcus Johnson'],
        # blank row 9
        ['MLS-10200', '432 Spyglass Rd', 'Austin', 4, 3, 2500, '$838,000', 'Lisa Tanaka'],
        ['MLS-10201', '2250 Exposition Blvd', 'Austin', 3, 2, 1870, '$598,000', 'David Kim'],
        ['MLS-10202', '955 Lake Austin Blvd', 'Austin', 5, 4, 3790, '$1.6M', 'Jennifer Walsh'],
        ['MLS-10203', '601 Riverside Dr', 'Austin', 4, 3, 2560, 860000, 'Sarah Chen'],
        ['MLS-10204', '1344 N Lamar Blvd', 'Austin', 3, 2, 1920, '$572,000', 'Robert Hawkins'],
        ['MLS-10205', '3300 Bee Cave Rd', 'Austin', 5, 4, 3870, '$1.45M', 'Marcus Johnson'],
        ['MLS-10206', '124 Great Hills Trl', 'Austin', 4, 3, 2310, '$718,000', 'Lisa Tanaka'],
        # blank row 10
        ['MLS-10207', '780 Jollyville Rd', 'Austin', 3, 2, 1800, '$492,000', 'David Kim'],
        ['MLS-10208', '2120 W Anderson Ln', 'Austin', 4, 3, 2450, '$785,500', 'Jennifer Walsh'],
        ['MLS-10209', '515 W 38th St', 'Austin', 3, 2, 1680, '$615,000', 'Sarah Chen'],
        ['MLS-10210', '88 Music Ln', 'Austin', 4, 3, 2190, '$945,000', 'Robert Hawkins'],
        ['MLS-10211', '1050 Rainey St', 'Austin', 5, 4, 3100, '$1.75M', 'Marcus Johnson'],
    ]

    # Sentinel rows that mark blank row positions (index in 0-based listings_data where None=blank row)
    # We'll insert blank rows at strategic positions matching the data comments above
    blank_row_indices = {5, 13, 20, 27, 33, 40, 47, 54, 60, 67}  # 0-based indices in final row sequence

    # Build the final sequence with blanks inserted
    final_rows = []
    blank_positions_sorted = sorted(blank_row_indices)
    data_idx = 0
    blank_count = 0
    row_seq = 0  # tracks position in output (0-based)

    # We need to rebuild: insert blank rows at the right positions
    # Instead, let's just build the list directly:
    final_rows = []
    for i, item in enumerate(listings_data):
        if item is None:
            final_rows.append(None)
        else:
            final_rows.append(item)

    # Actually, let's rebuild from scratch with explicit None blanks
    final_sequence = []
    listing_iter = iter(listings_data)

    # Group 1: rows 0-4 (5 listings), then blank
    for _ in range(5):
        final_sequence.append(next(listing_iter))
    final_sequence.append(None)  # blank

    # Group 2: rows 5-11 (7 listings), then blank
    for _ in range(7):
        final_sequence.append(next(listing_iter))
    final_sequence.append(None)  # blank

    # Group 3: rows 12-19 (8 listings), then blank
    for _ in range(8):
        final_sequence.append(next(listing_iter))
    final_sequence.append(None)  # blank

    # Group 4: rows 20-26 (7 listings), then blank
    for _ in range(7):
        final_sequence.append(next(listing_iter))
    final_sequence.append(None)  # blank

    # Group 5: rows 27-32 (6 listings), then blank
    for _ in range(6):
        final_sequence.append(next(listing_iter))
    final_sequence.append(None)  # blank

    # Group 6: rows 33-39 (7 listings), then blank
    for _ in range(7):
        final_sequence.append(next(listing_iter))
    final_sequence.append(None)  # blank

    # Group 7: rows 40-46 (7 listings), then blank
    for _ in range(7):
        final_sequence.append(next(listing_iter))
    final_sequence.append(None)  # blank

    # Group 8: rows 47-53 (7 listings), then blank
    for _ in range(7):
        final_sequence.append(next(listing_iter))
    final_sequence.append(None)  # blank

    # Group 9: rows 54-59 (6 listings), then blank
    for _ in range(6):
        final_sequence.append(next(listing_iter))
    final_sequence.append(None)  # blank

    # Group 10: rows 60-66 (7 listings), then blank
    for _ in range(7):
        final_sequence.append(next(listing_iter))
    final_sequence.append(None)  # blank

    # Final group: remaining
    for item in listing_iter:
        final_sequence.append(item)

    # Write rows to the worksheet (starting at row 2)
    current_row = 2
    for item in final_sequence:
        if item is None:
            # blank row: just skip (leave row empty)
            current_row += 1
        else:
            for col, val in enumerate(item, 1):
                ws.cell(row=current_row, column=col, value=val)
            current_row += 1

    # --- Add merged cells in various locations ---
    # Merge some address+city cells (mimicking exported data with merged headers or grouped regions)
    # Merge a few address cells across B and C columns for some rows (region groupings)
    # Row 3-4: merge city column for Cedar Park / Round Rock area indicator
    # Row 8-9: merge address area
    # Row 20-21: merge MLS and address for same building
    # Row 33-34: merge city for same area

    # Find actual row numbers for specific MLS entries
    # MLS-10144 (Round Rock) is at row 4, MLS-10145 (Cedar Park) at row 5
    # Let's merge B4:C4 (address+city combined for a row)
    ws.merge_cells('B4:C4')
    ws['B4'] = '112 Lakewood Ct, Round Rock'
    ws['D4'] = 5  # restore beds

    # Merge B8:C8 for another listing
    ws.merge_cells('B8:C8')
    ws['B8'] = '1480 Highland Hills Dr, Austin'
    ws['D8'] = 6  # restore beds

    # Merge B15:C15
    ws.merge_cells('B15:C15')
    ws['B15'] = '88 Magnolia Park Ct, Leander'
    ws['D15'] = 4

    # Merge B23:C23
    ws.merge_cells('B23:C23')
    ws['B23'] = '756 Balcones Dr, Austin'
    ws['D23'] = 5

    # Merge B29:C29
    ws.merge_cells('B29:C29')
    ws['B29'] = '2677 Chisholm Trail, Georgetown'
    ws['D29'] = 3

    # Merge B36:C36
    ws.merge_cells('B36:C36')
    ws['B36'] = '3322 Bull Creek Rd, Austin'
    ws['D36'] = 5

    # Merge B44:C44
    ws.merge_cells('B44:C44')
    ws['B44'] = '1760 Reunion Ranch Rd, Georgetown'
    ws['D44'] = 5

    # Merge B50:C50
    ws.merge_cells('B50:C50')
    ws['B50'] = '2980 Lime Rock Dr, Leander'
    ws['D50'] = 4

    # Merge B56:C56
    ws.merge_cells('B56:C56')
    ws['B56'] = '1567 Old Spicewood Rd, Leander'
    ws['D56'] = 3

    # Merge B63:C63
    ws.merge_cells('B63:C63')
    ws['B63'] = '2250 Exposition Blvd, Austin'
    ws['D63'] = 3

    # Merge B68:C68
    ws.merge_cells('B68:C68')
    ws['B68'] = '780 Jollyville Rd, Austin'
    ws['D68'] = 3

    # Additional: merge some MLS# cells across two rows to simulate grouped data
    # Row 38-39 MLS# merged
    ws.merge_cells('A71:A72')
    ws['A71'] = 'MLS-10205'

    # Row 74-75 MLS# merged
    ws.merge_cells('A75:A76')
    ws['A75'] = 'MLS-10209'

    # Wider merge: B78:C78 for last group
    ws.merge_cells('B78:C78')
    ws['B78'] = '88 Music Ln, Austin'
    ws['D78'] = 4

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet: Listings')
    print(f'Rows written: {current_row - 2} (including blank rows)')
    print(f'Merged cell ranges: B4:C4, B8:C8, B15:C15, B23:C23, B29:C29, B36:C36, B44:C44, B50:C50, B56:C56, B63:C63, B68:C68, A71:A72, A75:A76, B78:C78')


create_initial()
