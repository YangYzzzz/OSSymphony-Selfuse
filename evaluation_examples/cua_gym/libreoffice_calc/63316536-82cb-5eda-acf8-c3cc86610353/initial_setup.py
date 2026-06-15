"""
Initial Setup: Freeze header row in regional sheets
Task ID: calc_sht_freeze_row_003
Domain: libreoffice_calc

Creates a workbook with six sheets:
- Global Summary: Aggregate view
- APAC, EMEA, AMER, LATAM, MEA: Regional sheets with 80-150 rows each
None of the sheets have freeze panes (the task is to add them).
"""

import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

random.seed(42)

WORKDIR = '/home/user'
TASK_ID = 'calc_sht_freeze_row_003'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

# Regional data pools
APAC_COUNTRIES = [
    'China', 'Japan', 'South Korea', 'India', 'Australia',
    'Singapore', 'Thailand', 'Vietnam', 'Indonesia', 'Malaysia',
    'Philippines', 'New Zealand', 'Taiwan', 'Hong Kong', 'Bangladesh'
]
EMEA_COUNTRIES = [
    'Germany', 'France', 'United Kingdom', 'Italy', 'Spain',
    'Netherlands', 'Sweden', 'Switzerland', 'Norway', 'Denmark',
    'Poland', 'Belgium', 'Austria', 'Finland', 'South Africa',
    'Nigeria', 'Egypt', 'Kenya', 'UAE', 'Saudi Arabia'
]
AMER_COUNTRIES = [
    'United States', 'Canada', 'Mexico', 'Brazil', 'Argentina',
    'Colombia', 'Chile', 'Peru', 'Venezuela', 'Ecuador'
]
LATAM_COUNTRIES = [
    'Brazil', 'Mexico', 'Argentina', 'Colombia', 'Chile',
    'Peru', 'Venezuela', 'Ecuador', 'Bolivia', 'Paraguay',
    'Uruguay', 'Guatemala', 'Honduras', 'El Salvador', 'Costa Rica'
]
MEA_COUNTRIES = [
    'Saudi Arabia', 'UAE', 'Qatar', 'Kuwait', 'Bahrain',
    'Oman', 'Jordan', 'Lebanon', 'Egypt', 'Morocco',
    'Nigeria', 'South Africa', 'Kenya', 'Ghana', 'Ethiopia'
]

PRODUCTS = [
    'Enterprise Suite', 'Cloud Storage Pro', 'Analytics Platform',
    'Security Shield', 'DevOps Toolkit', 'Data Warehouse',
    'AI Assistant', 'CRM System', 'ERP Solution', 'Marketing Hub',
    'HR Management', 'Finance Module', 'Supply Chain Pro', 'IoT Gateway',
    'Collaboration Tools'
]


def make_regional_rows(region_name, countries, num_rows):
    """Generate realistic sales data rows for a regional sheet."""
    rows = []
    for i in range(num_rows):
        country = random.choice(countries)
        product = random.choice(PRODUCTS)
        q1 = round(random.uniform(12000, 980000), 2)
        q2 = round(q1 * random.uniform(0.75, 1.35), 2)
        q3 = round(q1 * random.uniform(0.80, 1.40), 2)
        q4 = round(q1 * random.uniform(0.85, 1.50), 2)
        annual = round(q1 + q2 + q3 + q4, 2)
        rows.append([region_name, country, product, q1, q2, q3, q4, annual])
    return rows


def add_regional_sheet(wb, sheet_name, region_name, countries, num_rows):
    """Add a regional sheet with header row and data, NO freeze panes."""
    ws = wb.create_sheet(sheet_name)
    headers = ['Region', 'Country', 'Product', 'Q1', 'Q2', 'Q3', 'Q4', 'Annual']

    # Write headers (plain, unbolded - just data)
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Write data rows
    data_rows = make_regional_rows(region_name, countries, num_rows)
    for r, row_data in enumerate(data_rows, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 22
    for col_letter in ['D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col_letter].width = 14

    # Number format for financial columns
    for r in range(2, num_rows + 2):
        for c in range(4, 9):
            ws.cell(row=r, column=c).number_format = '#,##0.00'

    # NO freeze panes — that is the task to add


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Global Summary ---
    ws_global = wb.active
    ws_global.title = 'Global Summary'

    global_headers = ['Region', 'Total Countries', 'Total Products',
                      'Q1 Total', 'Q2 Total', 'Q3 Total', 'Q4 Total', 'Annual Total']
    for col, h in enumerate(global_headers, 1):
        ws_global.cell(row=1, column=col, value=h)

    global_data = [
        ['APAC', 15, 15, 8_453_200.50, 9_102_440.75, 8_876_300.20, 10_234_500.80, 36_666_442.25],
        ['EMEA', 20, 15, 12_304_800.00, 11_987_600.50, 13_450_200.75, 14_123_000.25, 51_865_601.50],
        ['AMER', 10, 15, 18_234_500.75, 19_876_300.00, 17_654_800.50, 21_345_600.25, 77_111_201.50],
        ['LATAM', 15, 15, 4_532_100.25, 4_876_500.75, 5_123_400.50, 5_789_300.00, 20_321_301.50],
        ['MEA',  15, 15, 3_456_700.50, 3_789_200.25, 4_023_500.75, 4_567_800.00, 15_837_201.50],
    ]
    for r, row_data in enumerate(global_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_global.cell(row=r, column=c, value=val)

    # Totals row
    ws_global.cell(row=7, column=1, value='TOTAL')
    for c in range(4, 9):
        col_letter = ['D', 'E', 'F', 'G', 'H'][c - 4]
        ws_global.cell(row=7, column=c,
                       value=f'=SUM({col_letter}2:{col_letter}6)')

    # Column widths for Global Summary
    ws_global.column_dimensions['A'].width = 14
    ws_global.column_dimensions['B'].width = 18
    ws_global.column_dimensions['C'].width = 18
    for col_letter in ['D', 'E', 'F', 'G', 'H']:
        ws_global.column_dimensions[col_letter].width = 16

    # --- Regional Sheets (no freeze panes) ---
    add_regional_sheet(wb, 'APAC',  'APAC',  APAC_COUNTRIES,  100)
    add_regional_sheet(wb, 'EMEA',  'EMEA',  EMEA_COUNTRIES,  120)
    add_regional_sheet(wb, 'AMER',  'AMER',  AMER_COUNTRIES,   90)
    add_regional_sheet(wb, 'LATAM', 'LATAM', LATAM_COUNTRIES,  85)
    add_regional_sheet(wb, 'MEA',   'MEA',   MEA_COUNTRIES,    80)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheets: {wb.sheetnames}')
    for sh_name in ['APAC', 'EMEA', 'AMER', 'LATAM', 'MEA']:
        ws = wb[sh_name]
        print(f'  {sh_name}: {ws.max_row - 1} data rows, freeze_panes={ws.freeze_panes!r}')


create_initial()
