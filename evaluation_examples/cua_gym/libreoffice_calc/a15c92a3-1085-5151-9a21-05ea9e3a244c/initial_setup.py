"""
Initial Setup: Competitive analysis tracker with deal data
Task ID: calc_sales_070
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_070'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: CompAnalysis ---
    ws1 = wb.active
    ws1.title = 'CompAnalysis'

    # Headers
    headers = ['Deal', 'Value', 'Result', 'Competitor']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # Data rows (8 deals)
    data = [
        ['D1', 150000, 'Won', 'CompX'],
        ['D2', 80000, 'Lost', 'CompY'],
        ['D3', 200000, 'Lost', 'CompX'],
        ['D4', 120000, 'Won', 'CompY'],
        ['D5', 300000, 'Lost', 'CompZ'],
        ['D6', 95000, 'Won', 'CompX'],
        ['D7', 175000, 'Lost', 'CompY'],
        ['D8', 250000, 'Won', 'CompZ'],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # --- Sheet 2: Summary (headers and competitor names only, NO formulas) ---
    ws2 = wb.create_sheet('Summary')

    summary_headers = ['Competitor', 'Encounters', 'Wins', 'Losses', 'Win Rate', 'Avg Loss Value']
    for col, h in enumerate(summary_headers, 1):
        ws2.cell(row=1, column=col, value=h)

    # Competitor names in A2:A4
    ws2.cell(row=2, column=1, value='CompX')
    ws2.cell(row=3, column=1, value='CompY')
    ws2.cell(row=4, column=1, value='CompZ')

    # Label in A6
    ws2.cell(row=6, column=1, value='Most Dangerous')

    # B2:F4 and B6 are intentionally left empty — task is to fill them

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
