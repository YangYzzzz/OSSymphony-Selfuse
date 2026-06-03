"""
Reward Script: Competitive analysis tracker with win rate, avg loss value, most dangerous competitor
Task ID: calc_sales_070
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Encounters formulas (COUNTIFS) in B2:B4
  Component 2 (0.20): Wins formulas (COUNTIFS with "Won") in C2:C4
  Component 3 (0.15): Losses formulas (COUNTIFS with "Lost") in D2:D4
  Component 4 (0.20): Win Rate formulas in E2:E4 (ratio C/B)
  Component 5 (0.10): Avg Loss Value formulas (AVERAGEIFS) in F2:F4
  Component 6 (0.10): Most Dangerous competitor formula in B6
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_070'


def normalize_formula(f):
    """Normalize a formula string for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def is_countifs_for_competitor(formula, competitor_ref, extra_criteria=None):
    """
    Check if formula is a COUNTIFS that counts rows matching a competitor.
    competitor_ref: e.g. 'A2' — the cell referencing the competitor name.
    extra_criteria: optional string like '"WON"' or '"LOST"' that must appear.
    """
    nf = normalize_formula(formula)
    if not nf.startswith('=COUNTIFS(') and not nf.startswith('=COUNTIF('):
        return False
    # Must reference the CompAnalysis D column (competitor column)
    if 'COMPANALYSIS!D' not in nf and 'COMPANALYSIS!$D' not in nf:
        return False
    # Must reference the competitor cell
    if competitor_ref.upper() not in nf:
        return False
    if extra_criteria:
        if extra_criteria.upper() not in nf:
            return False
    return True


def is_win_rate_formula(formula):
    """Check if formula computes win rate as ratio of wins to encounters (C/B in same row)."""
    nf = normalize_formula(formula)
    if not nf.startswith('='):
        return False
    # Accept patterns like =C2/B2 or =C2*100/B2 or =ROUND(C2/B2,...) etc.
    # Core requirement: references both C and B columns in the same row
    if re.search(r'C\d+', nf) and re.search(r'B\d+', nf):
        return True
    # Also accept COUNTIFS-based inline formulas
    if 'COUNTIFS(' in nf and 'WON' in nf:
        return True
    return False


def is_averageifs_formula(formula, competitor_ref):
    """Check if formula uses AVERAGEIFS to compute avg loss value for a competitor."""
    nf = normalize_formula(formula)
    if 'AVERAGEIFS(' not in nf and 'AVERAGEIF(' not in nf:
        return False
    # Must reference CompAnalysis sheet
    if 'COMPANALYSIS!' not in nf:
        return False
    # Must reference the competitor cell
    if competitor_ref.upper() not in nf:
        return False
    # Must reference "Lost" criterion
    if '"LOST"' not in nf:
        return False
    return True


def is_most_dangerous_formula(formula):
    """Check if B6 contains a formula that identifies the competitor with lowest win rate."""
    nf = normalize_formula(formula)
    if not nf.startswith('='):
        return False
    # Accept INDEX/MATCH pattern, or VLOOKUP, or other lookup approaches
    # The key is it should reference the win rate range (E column) and competitor range (A column)
    if 'INDEX(' in nf and 'MATCH(' in nf:
        return True
    if 'VLOOKUP(' in nf:
        return True
    # Also accept IF-based approaches
    if 'IF(' in nf and 'MIN(' in nf:
        return True
    return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Summary sheet must exist
    if 'Summary' not in wb.sheetnames:
        print("FAIL: Summary sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Summary']

    # Component 1: Encounters formulas (COUNTIFS) in B2:B4 (0.25 points)
    try:
        comp1_pass = 0
        for row_num, comp_ref in [(2, 'A2'), (3, 'A3'), (4, 'A4')]:
            cell_val = ws.cell(row=row_num, column=2).value  # B column
            if is_countifs_for_competitor(cell_val, comp_ref):
                comp1_pass += 1
                print(f"  PASS: B{row_num} has COUNTIFS for encounters: {cell_val}")
            else:
                print(f"  FAIL: B{row_num} expected COUNTIFS encounter formula, found: {cell_val!r}")
        if comp1_pass == 3:
            print(f"PASS: Component 1 — All 3 Encounters formulas correct (0.25 pts)")
            total_score += 0.25
        elif comp1_pass > 0:
            partial = round(0.25 * comp1_pass / 3, 4)
            print(f"PARTIAL: Component 1 — {comp1_pass}/3 Encounters formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No valid Encounters formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Wins formulas (COUNTIFS with "Won") in C2:C4 (0.20 points)
    try:
        comp2_pass = 0
        for row_num, comp_ref in [(2, 'A2'), (3, 'A3'), (4, 'A4')]:
            cell_val = ws.cell(row=row_num, column=3).value  # C column
            if is_countifs_for_competitor(cell_val, comp_ref, extra_criteria='"Won"'):
                comp2_pass += 1
                print(f"  PASS: C{row_num} has COUNTIFS for wins: {cell_val}")
            else:
                print(f"  FAIL: C{row_num} expected COUNTIFS wins formula, found: {cell_val!r}")
        if comp2_pass == 3:
            print(f"PASS: Component 2 — All 3 Wins formulas correct (0.20 pts)")
            total_score += 0.20
        elif comp2_pass > 0:
            partial = round(0.20 * comp2_pass / 3, 4)
            print(f"PARTIAL: Component 2 — {comp2_pass}/3 Wins formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No valid Wins formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Losses formulas (COUNTIFS with "Lost") in D2:D4 (0.15 points)
    try:
        comp3_pass = 0
        for row_num, comp_ref in [(2, 'A2'), (3, 'A3'), (4, 'A4')]:
            cell_val = ws.cell(row=row_num, column=4).value  # D column
            if is_countifs_for_competitor(cell_val, comp_ref, extra_criteria='"Lost"'):
                comp3_pass += 1
                print(f"  PASS: D{row_num} has COUNTIFS for losses: {cell_val}")
            else:
                print(f"  FAIL: D{row_num} expected COUNTIFS losses formula, found: {cell_val!r}")
        if comp3_pass == 3:
            print(f"PASS: Component 3 — All 3 Losses formulas correct (0.15 pts)")
            total_score += 0.15
        elif comp3_pass > 0:
            partial = round(0.15 * comp3_pass / 3, 4)
            print(f"PARTIAL: Component 3 — {comp3_pass}/3 Losses formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No valid Losses formulas found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Win Rate formulas in E2:E4 (0.20 points)
    try:
        comp4_pass = 0
        for row_num in [2, 3, 4]:
            cell_val = ws.cell(row=row_num, column=5).value  # E column
            if is_win_rate_formula(cell_val):
                comp4_pass += 1
                print(f"  PASS: E{row_num} has win rate formula: {cell_val}")
            else:
                print(f"  FAIL: E{row_num} expected win rate formula, found: {cell_val!r}")
        if comp4_pass == 3:
            print(f"PASS: Component 4 — All 3 Win Rate formulas correct (0.20 pts)")
            total_score += 0.20
        elif comp4_pass > 0:
            partial = round(0.20 * comp4_pass / 3, 4)
            print(f"PARTIAL: Component 4 — {comp4_pass}/3 Win Rate formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — No valid Win Rate formulas found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Avg Loss Value formulas (AVERAGEIFS) in F2:F4 (0.10 points)
    try:
        comp5_pass = 0
        for row_num, comp_ref in [(2, 'A2'), (3, 'A3'), (4, 'A4')]:
            cell_val = ws.cell(row=row_num, column=6).value  # F column
            if is_averageifs_formula(cell_val, comp_ref):
                comp5_pass += 1
                print(f"  PASS: F{row_num} has AVERAGEIFS formula: {cell_val}")
            else:
                print(f"  FAIL: F{row_num} expected AVERAGEIFS formula, found: {cell_val!r}")
        if comp5_pass == 3:
            print(f"PASS: Component 5 — All 3 Avg Loss Value formulas correct (0.10 pts)")
            total_score += 0.10
        elif comp5_pass > 0:
            partial = round(0.10 * comp5_pass / 3, 4)
            print(f"PARTIAL: Component 5 — {comp5_pass}/3 Avg Loss Value formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — No valid Avg Loss Value formulas found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Most Dangerous competitor formula in B6 (0.10 points)
    try:
        cell_val = ws.cell(row=6, column=2).value  # B6
        if is_most_dangerous_formula(cell_val):
            print(f"PASS: Component 6 — B6 has Most Dangerous formula: {cell_val} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 6 — B6 expected lookup formula for most dangerous competitor, found: {cell_val!r}")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
