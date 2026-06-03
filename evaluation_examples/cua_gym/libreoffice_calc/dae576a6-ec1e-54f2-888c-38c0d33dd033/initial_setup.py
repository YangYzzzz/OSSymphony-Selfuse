"""
Initial Setup: Cost Variance Analysis Report
Task ID: calc_ops_cost_analysis_variance_040
Domain: libreoffice_calc

Creates initial spreadsheet with CostVariance sheet containing:
- Headers in row 1 (A-K)
- 10 cost elements in rows 2-11 with Standard Qty/Rate and Actual Qty/Rate filled
- Columns D (Standard Cost), G (Actual Cost), H (Price Variance),
  I (Efficiency Variance), J (Total Variance), K (Variance %) left EMPTY
  for agent to fill with formulas
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_cost_analysis_variance_040'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'CostVariance'

    # --- Headers Row 1 ---
    headers = [
        'Cost Element',       # A
        'Standard Qty',       # B
        'Standard Rate',      # C
        'Standard Cost',      # D - empty, agent fills with =B*C
        'Actual Qty',         # E
        'Actual Rate',        # F
        'Actual Cost',        # G - empty, agent fills with =E*F
        'Price Variance',     # H - empty, agent fills with =(F-C)*E
        'Efficiency Variance',# I - empty, agent fills with =(E-B)*C
        'Total Variance',     # J - empty, agent fills with =G-D
        'Variance %',         # K - empty, agent fills with =J/D
    ]
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF', name='Calibri', size=11)
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    ws.row_dimensions[1].height = 30

    # --- Cost Elements Data (rows 2-11) ---
    # Columns: A (Cost Element), B (Std Qty), C (Std Rate), D (empty),
    #          E (Actual Qty), F (Actual Rate), G-K (empty)
    # Data represents last month's production cost elements
    cost_data = [
        # Cost Element,           Std Qty, Std Rate,  Act Qty, Act Rate
        ('Direct Labor - Assembly',     480,    22.50,     495,    23.10),
        ('Direct Labor - Machining',    320,    28.75,     310,    29.50),
        ('Raw Materials - Steel',      1200,     4.80,    1250,     4.65),
        ('Raw Materials - Aluminum',    850,     7.20,     840,     7.35),
        ('Raw Materials - Plastics',    600,     2.15,     625,     2.10),
        ('Raw Materials - Electronics', 400,    12.40,     395,    13.00),
        ('Raw Materials - Packaging',   950,     1.85,     960,     1.82),
        ('Variable Overhead',          1200,     6.50,    1200,     6.75),
        ('Fixed Overhead',             1200,    18.00,    1200,    18.00),
        ('Quality Control',             240,    15.00,     252,    14.75),
    ]

    data_font = Font(name='Calibri', size=11)
    alt_fill = PatternFill(start_color='FFE9EFF7', end_color='FFE9EFF7', fill_type='solid')

    for r_idx, row_data in enumerate(cost_data, 2):
        cost_elem, std_qty, std_rate, act_qty, act_rate = row_data
        row_fill = alt_fill if r_idx % 2 == 0 else None

        # A: Cost Element
        cell_a = ws.cell(row=r_idx, column=1, value=cost_elem)
        cell_a.font = data_font
        cell_a.border = border
        if row_fill:
            cell_a.fill = row_fill

        # B: Standard Qty
        cell_b = ws.cell(row=r_idx, column=2, value=std_qty)
        cell_b.font = data_font
        cell_b.alignment = Alignment(horizontal='right')
        cell_b.number_format = '#,##0'
        cell_b.border = border
        if row_fill:
            cell_b.fill = row_fill

        # C: Standard Rate
        cell_c = ws.cell(row=r_idx, column=3, value=std_rate)
        cell_c.font = data_font
        cell_c.alignment = Alignment(horizontal='right')
        cell_c.number_format = '$#,##0.00'
        cell_c.border = border
        if row_fill:
            cell_c.fill = row_fill

        # D: Standard Cost - EMPTY (agent fills formula)
        cell_d = ws.cell(row=r_idx, column=4, value=None)
        cell_d.border = border
        if row_fill:
            cell_d.fill = row_fill

        # E: Actual Qty
        cell_e = ws.cell(row=r_idx, column=5, value=act_qty)
        cell_e.font = data_font
        cell_e.alignment = Alignment(horizontal='right')
        cell_e.number_format = '#,##0'
        cell_e.border = border
        if row_fill:
            cell_e.fill = row_fill

        # F: Actual Rate
        cell_f = ws.cell(row=r_idx, column=6, value=act_rate)
        cell_f.font = data_font
        cell_f.alignment = Alignment(horizontal='right')
        cell_f.number_format = '$#,##0.00'
        cell_f.border = border
        if row_fill:
            cell_f.fill = row_fill

        # G-K: EMPTY (agent fills formulas)
        for col in range(7, 12):
            cell = ws.cell(row=r_idx, column=col, value=None)
            cell.border = border
            if row_fill:
                cell.fill = row_fill

    # --- Column widths ---
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 14
    ws.column_dimensions['K'].width = 12

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet: CostVariance')
    print(f'Rows: 1 header + 10 data rows (rows 2-11)')
    print(f'Filled columns: A (Cost Element), B (Std Qty), C (Std Rate), E (Act Qty), F (Act Rate)')
    print(f'Empty columns: D, G, H, I, J, K (agent fills with formulas)')


create_initial()
