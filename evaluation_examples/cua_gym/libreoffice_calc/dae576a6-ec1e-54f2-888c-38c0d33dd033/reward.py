"""
Reward Script: Cost Variance Analysis Report
Task ID: calc_ops_cost_analysis_variance_040
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Standard Cost formulas D2:D11 (=Bx*Cx)           — 0.25 pts
  Component 2: Actual Cost formulas G2:G11 (=Ex*Fx)             — 0.15 pts
  Component 3: Price Variance formulas H2:H11 (=(Fx-Cx)*Ex)     — 0.15 pts
  Component 4: Efficiency Variance formulas I2:I11 (=(Ex-Bx)*Cx)— 0.15 pts
  Component 5: Total Variance J2:J11 (=Gx-Dx) and
               Variance% K2:K11 (=Jx/Dx)                        — 0.10 pts
  Component 6: Total row 12 with SUM formulas D,G,H,I,J          — 0.10 pts
  Component 7: Conditional formatting on K2:K11                  — 0.10 pts
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_cost_analysis_variance_040'


def normalize_formula(f):
    """Normalize formula: uppercase, remove spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: CostVariance sheet must exist
    if 'CostVariance' not in wb.sheetnames:
        print("CRITICAL: 'CostVariance' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['CostVariance']

    # -----------------------------------------------------------------------
    # Component 1: Standard Cost formulas in D2:D11 (=Bx*Cx)  — 0.25 points
    # Each correct formula earns 0.025 points (10 rows)
    # -----------------------------------------------------------------------
    try:
        std_cost_count = 0
        for row in range(2, 12):
            val = ws.cell(row=row, column=4).value  # Column D
            if isinstance(val, str):
                norm = normalize_formula(val)
                expected = normalize_formula(f'=B{row}*C{row}')
                if norm == expected:
                    std_cost_count += 1
        if std_cost_count == 10:
            print(f"PASS: Component 1 — All 10 Standard Cost formulas D2:D11 correct (0.25 pts)")
            total_score += 0.25
        elif std_cost_count >= 5:
            partial = round(0.025 * std_cost_count, 4)
            print(f"PARTIAL: Component 1 — {std_cost_count}/10 Standard Cost formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {std_cost_count}/10 Standard Cost formulas (D2:D11) present; expected =Bx*Cx")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Actual Cost formulas in G2:G11 (=Ex*Fx)  — 0.15 points
    # -----------------------------------------------------------------------
    try:
        act_cost_count = 0
        for row in range(2, 12):
            val = ws.cell(row=row, column=7).value  # Column G
            if isinstance(val, str):
                norm = normalize_formula(val)
                expected = normalize_formula(f'=E{row}*F{row}')
                if norm == expected:
                    act_cost_count += 1
        if act_cost_count == 10:
            print(f"PASS: Component 2 — All 10 Actual Cost formulas G2:G11 correct (0.15 pts)")
            total_score += 0.15
        elif act_cost_count >= 5:
            partial = round(0.015 * act_cost_count, 4)
            print(f"PARTIAL: Component 2 — {act_cost_count}/10 Actual Cost formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {act_cost_count}/10 Actual Cost formulas (G2:G11) present; expected =Ex*Fx")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Price Variance formulas H2:H11 (=(Fx-Cx)*Ex)  — 0.15 points
    # -----------------------------------------------------------------------
    try:
        price_var_count = 0
        for row in range(2, 12):
            val = ws.cell(row=row, column=8).value  # Column H
            if isinstance(val, str):
                norm = normalize_formula(val)
                expected = normalize_formula(f'=(F{row}-C{row})*E{row}')
                if norm == expected:
                    price_var_count += 1
        if price_var_count == 10:
            print(f"PASS: Component 3 — All 10 Price Variance formulas H2:H11 correct (0.15 pts)")
            total_score += 0.15
        elif price_var_count >= 5:
            partial = round(0.015 * price_var_count, 4)
            print(f"PARTIAL: Component 3 — {price_var_count}/10 Price Variance formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {price_var_count}/10 Price Variance formulas (H2:H11); expected =(Fx-Cx)*Ex")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: Efficiency Variance formulas I2:I11 (=(Ex-Bx)*Cx)  — 0.15 points
    # -----------------------------------------------------------------------
    try:
        eff_var_count = 0
        for row in range(2, 12):
            val = ws.cell(row=row, column=9).value  # Column I
            if isinstance(val, str):
                norm = normalize_formula(val)
                expected = normalize_formula(f'=(E{row}-B{row})*C{row}')
                if norm == expected:
                    eff_var_count += 1
        if eff_var_count == 10:
            print(f"PASS: Component 4 — All 10 Efficiency Variance formulas I2:I11 correct (0.15 pts)")
            total_score += 0.15
        elif eff_var_count >= 5:
            partial = round(0.015 * eff_var_count, 4)
            print(f"PARTIAL: Component 4 — {eff_var_count}/10 Efficiency Variance formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {eff_var_count}/10 Efficiency Variance formulas (I2:I11); expected =(Ex-Bx)*Cx")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -----------------------------------------------------------------------
    # Component 5: Total Variance J2:J11 (=Gx-Dx) and Variance% K2:K11 (=Jx/Dx)
    # — 0.10 points (0.05 each)
    # -----------------------------------------------------------------------
    try:
        total_var_count = 0
        for row in range(2, 12):
            val = ws.cell(row=row, column=10).value  # Column J
            if isinstance(val, str):
                norm = normalize_formula(val)
                expected = normalize_formula(f'=G{row}-D{row}')
                if norm == expected:
                    total_var_count += 1

        var_pct_count = 0
        for row in range(2, 12):
            val = ws.cell(row=row, column=11).value  # Column K
            if isinstance(val, str):
                norm = normalize_formula(val)
                expected = normalize_formula(f'=J{row}/D{row}')
                if norm == expected:
                    var_pct_count += 1

        if total_var_count == 10:
            print(f"PASS: Component 5a — All 10 Total Variance formulas J2:J11 correct (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5a — Only {total_var_count}/10 Total Variance formulas (J2:J11); expected =Gx-Dx")

        if var_pct_count == 10:
            print(f"PASS: Component 5b — All 10 Variance% formulas K2:K11 correct (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5b — Only {var_pct_count}/10 Variance% formulas (K2:K11); expected =Jx/Dx")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # -----------------------------------------------------------------------
    # Component 6: Total row 12 with SUM formulas for D,G,H,I,J  — 0.10 points
    # -----------------------------------------------------------------------
    try:
        # Check A12 has 'Total' label and D12,G12,H12,I12,J12 have SUM formulas
        a12 = ws.cell(row=12, column=1).value
        label_ok = a12 is not None and str(a12).strip().upper() == 'TOTAL'

        sum_cols = {4: 'D', 7: 'G', 8: 'H', 9: 'I', 10: 'J'}
        sum_count = 0
        for col_idx, col_letter in sum_cols.items():
            val = ws.cell(row=12, column=col_idx).value
            if isinstance(val, str):
                norm = normalize_formula(val)
                expected = normalize_formula(f'=SUM({col_letter}2:{col_letter}11)')
                if norm == expected:
                    sum_count += 1

        if label_ok and sum_count == 5:
            print(f"PASS: Component 6 — Total row 12 with label and all 5 SUM formulas (0.10 pts)")
            total_score += 0.10
        elif sum_count >= 3:
            partial = round(0.02 * sum_count, 4)
            print(f"PARTIAL: Component 6 — Total row: label={'OK' if label_ok else 'MISSING'}, {sum_count}/5 SUM formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 6 — Total row 12 missing; label={'OK' if label_ok else 'MISSING'}, only {sum_count}/5 SUM formulas present")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # -----------------------------------------------------------------------
    # Component 7: Conditional formatting on K2:K11  — 0.10 points
    # Expects: expression formula for ABS(K)>5% (red fill) and ABS(K)<=5% (green fill)
    # -----------------------------------------------------------------------
    try:
        cf_list = list(ws.conditional_formatting)
        # Look for conditional formatting covering K column (column 11)
        k_cf_ranges = []
        for cf in cf_list:
            cf_str = str(cf)
            if 'K' in cf_str.upper():
                k_cf_ranges.append(cf)

        if not k_cf_ranges:
            print("FAIL: Component 7 — No conditional formatting found on K column (K2:K11)")
        else:
            # Check for at least 2 expression rules covering ±5% threshold
            rules_found = []
            for cf in k_cf_ranges:
                for rule in cf.rules:
                    if rule.type == 'expression' and hasattr(rule, 'formula') and rule.formula:
                        formula_str = str(rule.formula[0]).upper().replace(' ', '')
                        rules_found.append(formula_str)

            has_red_rule = any('ABS' in r and '0.05' in r for r in rules_found)
            has_green_rule = any('ABS' in r and '0.05' in r for r in rules_found)
            # We need at least 2 rules: one for >5% and one for <=5%
            has_both_rules = len(rules_found) >= 2

            if has_both_rules:
                print(f"PASS: Component 7 — Conditional formatting on K column with {len(rules_found)} expression rules (0.10 pts)")
                total_score += 0.10
            elif len(rules_found) == 1:
                print(f"PARTIAL: Component 7 — Only 1 conditional formatting rule on K column (0.05 pts); expected 2")
                total_score += 0.05
            else:
                print(f"FAIL: Component 7 — Conditional formatting on K column exists but no expression rules found")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
