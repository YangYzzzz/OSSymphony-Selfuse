"""
Initial Setup: Financial report with columns in scrambled order
Task ID: osworld_calc_reorder_columns_005
Domain: libreoffice_calc

Creates a spreadsheet with the financial report columns in the WRONG order:
  Debit, Account Name, Note, Credit, Account Code, Closing Balance, Opening Balance
The task asks the agent to reorder them to:
  Account Code, Account Name, Opening Balance, Debit, Credit, Closing Balance, Note
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_reorder_columns_005'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Financial Report ---
    ws = wb.active
    ws.title = "Financial Report"

    # Columns in SCRAMBLED order (this is what the agent must fix):
    # Debit, Account Name, Note, Credit, Account Code, Closing Balance, Opening Balance
    headers = [
        "Debit",
        "Account Name",
        "Note",
        "Credit",
        "Account Code",
        "Closing Balance",
        "Opening Balance",
    ]

    # Write header row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Realistic financial account data
    # Format per scrambled column order:
    # Debit | Account Name            | Note                    | Credit   | Account Code | Closing Balance | Opening Balance
    data = [
        [12500.00, "Cash and Cash Equivalents",  "Petty cash replenishment",  0.00,     "1001", 87350.00,  99850.00 ],
        [0.00,     "Accounts Receivable",         "Invoice #INV-20240315",    35000.00, "1100", 152000.00, 117000.00],
        [4800.00,  "Prepaid Insurance",           "Q2 insurance premium",     0.00,     "1200", 19200.00,  24000.00 ],
        [0.00,     "Office Equipment",            "",                          0.00,     "1500", 48500.00,  48500.00 ],
        [0.00,     "Accumulated Depreciation",    "Annual depreciation",      1500.00,  "1510", -9500.00,  -8000.00 ],
        [0.00,     "Accounts Payable",            "Supplier invoices due",    22300.00, "2001", -65400.00, -43100.00],
        [5000.00,  "Salaries Payable",            "March payroll accrual",    0.00,     "2100", -18200.00, -23200.00],
        [0.00,     "Bank Loan - Long Term",       "",                          0.00,     "2500", -120000.00,-120000.00],
        [0.00,     "Share Capital",               "Founding equity",          0.00,     "3001", -50000.00, -50000.00],
        [0.00,     "Retained Earnings",           "Prior year profit",        18450.00, "3100", -81350.00, -62900.00],
        [98600.00, "Sales Revenue",               "Product line A & B",       0.00,     "4001", -186700.00,-285300.00],
        [0.00,     "Service Revenue",             "Consulting engagements",   14200.00, "4100", -52400.00, -38200.00],
        [31200.00, "Cost of Goods Sold",          "Inventory cost allocated", 0.00,     "5001", 178900.00, 147700.00],
        [8750.00,  "Salaries Expense",            "Administrative staff",     0.00,     "6001", 96250.00,  87500.00 ],
        [1100.00,  "Utilities Expense",           "Electricity and water",    0.00,     "6100", 9800.00,   8700.00  ],
        [650.00,   "Office Supplies Expense",     "Stationery and consumables",0.00,    "6200", 3850.00,   3200.00  ],
        [2200.00,  "Marketing Expense",           "Online advertising Q1",    0.00,     "6300", 14600.00,  12400.00 ],
        [920.00,   "Travel and Entertainment",    "Client meetings",          0.00,     "6400", 5740.00,   4820.00  ],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths
    col_widths = [14, 30, 28, 14, 14, 18, 18]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
