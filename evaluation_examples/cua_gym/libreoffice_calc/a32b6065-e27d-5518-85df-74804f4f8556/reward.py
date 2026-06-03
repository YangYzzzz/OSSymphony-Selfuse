"""
Reward Script: Reorganize financial report columns to logical sequence
Task ID: osworld_calc_reorder_columns_005
Domain: libreoffice_calc
Scoring:
  Component 1: Column headers are in the correct order (0.5 pts)
  Component 2: Data values are in correct positional columns per the new order (0.3 pts)
  Component 3: All data rows are preserved with headers correct (0.2 pts)
Total: 1.0

NOTE on Component 2 design:
  We verify data values by FIXED column position (col 1=Account Code, col 3=Opening Balance, etc.)
  This FAILS on initial_env (values are in wrong positions) and PASSES on golden_env (correct positions).
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_reorder_columns_005'

# Expected column order after the task is complete (1-indexed)
EXPECTED_HEADERS = [
    'Account Code',       # col 1
    'Account Name',       # col 2
    'Opening Balance',    # col 3
    'Debit',              # col 4
    'Credit',             # col 5
    'Closing Balance',    # col 6
    'Note',               # col 7
]

# Expected data row count (18 data rows)
EXPECTED_DATA_ROWS = 18

# Known data rows in the CORRECT column order (col1=Account Code, col2=Account Name,
# col3=Opening Balance, col4=Debit, col5=Credit, col6=Closing Balance, col7=Note).
# These are the values that should appear at these FIXED column positions in the golden file.
# In the initial file, col1=Debit, col2=Account Name, col3=Note, col4=Credit,
# col5=Account Code, col6=Closing Balance, col7=Opening Balance
# so these same checks will FAIL on initial_env.
KNOWN_ROWS_BY_POSITION = [
    # (row_num, col1_account_code, col3_opening_balance, col4_debit, col5_credit, col6_closing_balance)
    (2,  '1001', 99850,    12500,  0,     87350),
    (3,  '1100', 117000,   0,      35000, 152000),
    (12, '4001', -285300,  98600,  0,     -186700),
    (19, '6400', 4820,     920,    0,     5740),
]


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get the Financial Report sheet
    try:
        if 'Financial Report' in wb.sheetnames:
            ws = wb['Financial Report']
        else:
            ws = wb.active
    except Exception as e:
        print(f"CRITICAL: Cannot access worksheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Column headers are in correct order (0.5 points)
    # FAILS on initial (headers in scrambled order), PASSES on golden (correct order)
    headers_correct = False
    try:
        actual_headers = [ws.cell(row=1, column=col).value for col in range(1, 8)]

        if actual_headers == EXPECTED_HEADERS:
            print(f"PASS: Component 1 — All 7 column headers in correct order (0.5 pts)")
            print(f"  Headers: {actual_headers}")
            total_score += 0.5
            headers_correct = True
        else:
            print(f"FAIL: Component 1 — Column header order incorrect.")
            print(f"  Expected: {EXPECTED_HEADERS}")
            print(f"  Actual:   {actual_headers}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Data values are at the correct column positions (0.3 points)
    # We check known data values at FIXED column positions (col 1, 3, 4, 5, 6).
    # Initial file: col1=Debit(12500), col3=Note(text), col4=Credit(0) for row 2
    #   → All position checks FAIL because values don't match the expected positional values
    # Golden file: col1=Account Code('1001'), col3=Opening Balance(99850), col4=Debit(12500)
    #   → All position checks PASS
    try:
        checks_passed = 0
        checks_total = len(KNOWN_ROWS_BY_POSITION)

        for row_num, exp_code, exp_ob, exp_dr, exp_cr, exp_cb in KNOWN_ROWS_BY_POSITION:
            row_ok = True

            # Check col 1 = Account Code
            actual_code = ws.cell(row=row_num, column=1).value
            if str(actual_code) != str(exp_code):
                print(f"FAIL: Component 2 — Row {row_num} col1 (Account Code): expected '{exp_code}', got '{actual_code}'")
                row_ok = False

            # Check col 3 = Opening Balance (numeric)
            try:
                actual_ob = float(ws.cell(row=row_num, column=3).value)
                if abs(actual_ob - exp_ob) > 0.01:
                    print(f"FAIL: Component 2 — Row {row_num} col3 (Opening Balance): expected {exp_ob}, got {actual_ob}")
                    row_ok = False
            except (TypeError, ValueError):
                print(f"FAIL: Component 2 — Row {row_num} col3 (Opening Balance) non-numeric: {ws.cell(row=row_num, column=3).value}")
                row_ok = False

            # Check col 4 = Debit (numeric)
            try:
                actual_dr = float(ws.cell(row=row_num, column=4).value)
                if abs(actual_dr - exp_dr) > 0.01:
                    print(f"FAIL: Component 2 — Row {row_num} col4 (Debit): expected {exp_dr}, got {actual_dr}")
                    row_ok = False
            except (TypeError, ValueError):
                print(f"FAIL: Component 2 — Row {row_num} col4 (Debit) non-numeric: {ws.cell(row=row_num, column=4).value}")
                row_ok = False

            # Check col 5 = Credit (numeric)
            try:
                actual_cr = float(ws.cell(row=row_num, column=5).value)
                if abs(actual_cr - exp_cr) > 0.01:
                    print(f"FAIL: Component 2 — Row {row_num} col5 (Credit): expected {exp_cr}, got {actual_cr}")
                    row_ok = False
            except (TypeError, ValueError):
                print(f"FAIL: Component 2 — Row {row_num} col5 (Credit) non-numeric: {ws.cell(row=row_num, column=5).value}")
                row_ok = False

            # Check col 6 = Closing Balance (numeric)
            try:
                actual_cb = float(ws.cell(row=row_num, column=6).value)
                if abs(actual_cb - exp_cb) > 0.01:
                    print(f"FAIL: Component 2 — Row {row_num} col6 (Closing Balance): expected {exp_cb}, got {actual_cb}")
                    row_ok = False
            except (TypeError, ValueError):
                print(f"FAIL: Component 2 — Row {row_num} col6 (Closing Balance) non-numeric: {ws.cell(row=row_num, column=6).value}")
                row_ok = False

            if row_ok:
                checks_passed += 1

        if checks_passed == checks_total:
            print(f"PASS: Component 2 — All {checks_passed}/{checks_total} positional data checks passed (0.3 pts)")
            total_score += 0.3
        elif checks_passed > 0:
            partial = round(0.3 * checks_passed / checks_total, 2)
            print(f"PARTIAL: Component 2 — {checks_passed}/{checks_total} positional data checks passed ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — 0/{checks_total} positional data checks passed")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: All 18 data rows preserved AND headers are correct (0.2 points)
    # Anchored to headers_correct so it does NOT pass on initial_env
    try:
        actual_data_rows = ws.max_row - 1  # subtract header row

        if headers_correct and actual_data_rows == EXPECTED_DATA_ROWS:
            print(f"PASS: Component 3 — All {actual_data_rows} data rows preserved with correct column order (0.2 pts)")
            total_score += 0.2
        elif not headers_correct:
            print(f"FAIL: Component 3 — Headers not in correct order; skipping row count award")
        else:
            print(f"FAIL: Component 3 — Expected {EXPECTED_DATA_ROWS} data rows, found {actual_data_rows}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the VM environment
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
