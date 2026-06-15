"""
Reward Script: Set up milestone tracking for product launch project
Task ID: calc_ops_project_tracking_milestone_013
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Dropdown data validation on D2:D13            — 0.25 pts
  Component 2: Days Remaining formulas (=Cx-TODAY()) E2:E13  — 0.25 pts
  Component 3: RAG Status IF/OR formulas in F2:F13           — 0.25 pts
  Component 4: Conditional formatting on A2:F13 (R/A/G fills)— 0.15 pts
  Component 5: Freeze row 1 (freeze_panes == "A2")           — 0.10 pts
  Total: 1.00
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_ops_project_tracking_milestone_013'


def normalize_formula(f):
    """Normalize formula for comparison: uppercase, strip whitespace."""
    if not isinstance(f, str):
        return ''
    return f.strip().upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Milestones' sheet must exist
    if 'Milestones' not in wb.sheetnames:
        print("FAIL: Sheet 'Milestones' not found in workbook")
        print(f"Score: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws = wb['Milestones']

    # -----------------------------------------------------------------------
    # Component 1: Data Validation dropdown on D2:D13 (0.25 points)
    # Expected: list type with "Not Started,In Progress,Complete,Blocked"
    # -----------------------------------------------------------------------
    try:
        validations = ws.data_validations.dataValidation
        matched_dv = None
        for dv in validations:
            if dv.type == 'list' and 'D2' in str(dv.sqref) and 'D13' in str(dv.sqref):
                matched_dv = dv
                break
        if matched_dv is not None:
            f1_upper = (matched_dv.formula1 or '').upper()
            required_values = ['NOT STARTED', 'IN PROGRESS', 'COMPLETE', 'BLOCKED']
            all_present = all(r in f1_upper for r in required_values)
            if all_present:
                print(f"PASS: Component 1 — Data validation dropdown on D2:D13 with correct values (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 1 — Dropdown found on D2:D13 but missing required values. formula1={matched_dv.formula1}")
        else:
            print(f"FAIL: Component 1 — No list data validation found covering D2:D13")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Days Remaining formulas in E2:E13 (0.25 points)
    # Expected: each cell contains formula =Cx-TODAY() (case-insensitive)
    # -----------------------------------------------------------------------
    try:
        days_formula_count = 0
        days_formula_errors = []
        for row in range(2, 14):
            cell_val = ws.cell(row=row, column=5).value  # column E = 5
            if isinstance(cell_val, str):
                norm = normalize_formula(cell_val)
                expected = normalize_formula(f'=C{row}-TODAY()')
                if norm == expected:
                    days_formula_count += 1
                else:
                    days_formula_errors.append(f"E{row}: expected '=C{row}-TODAY()', got '{cell_val}'")
            else:
                days_formula_errors.append(f"E{row}: expected formula, got {repr(cell_val)}")

        if days_formula_count == 12:
            print(f"PASS: Component 2 — All 12 Days Remaining formulas (=Cx-TODAY()) present in E2:E13 (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — Only {days_formula_count}/12 E-column formulas correct. Errors: {days_formula_errors[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: RAG Status IF/OR formulas in F2:F13 (0.25 points)
    # Expected: =IF(OR(Dx="Blocked",Ex<0),"Red",IF(Ex<=7,"Amber","Green"))
    # -----------------------------------------------------------------------
    try:
        rag_formula_count = 0
        rag_formula_errors = []
        for row in range(2, 14):
            cell_val = ws.cell(row=row, column=6).value  # column F = 6
            if isinstance(cell_val, str):
                norm = normalize_formula(cell_val)
                expected = normalize_formula(
                    f'=IF(OR(D{row}="Blocked",E{row}<0),"Red",IF(E{row}<=7,"Amber","Green"))'
                )
                if norm == expected:
                    rag_formula_count += 1
                else:
                    rag_formula_errors.append(f"F{row}: got '{cell_val}'")
            else:
                rag_formula_errors.append(f"F{row}: expected formula, got {repr(cell_val)}")

        if rag_formula_count == 12:
            print(f"PASS: Component 3 — All 12 RAG Status formulas correct in F2:F13 (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 3 — Only {rag_formula_count}/12 F-column formulas correct. Errors: {rag_formula_errors[:3]}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: Conditional Formatting on A2:F13 with RAG fills (0.15 pts)
    # Expected: 3 FormulaRule expression rules referencing $F column
    # with Red (FFFF0000), Amber/Yellow (FFFFFF00), Green (FF00FF00) fills
    # -----------------------------------------------------------------------
    try:
        cf_rules = ws.conditional_formatting

        # Collect all expression rules that reference $F column
        red_rules = [
            rule
            for cf_range in cf_rules
            for rule in cf_range.rules
            if rule.type == 'expression'
            and rule.formula
            and 'RED' in normalize_formula(rule.formula[0])
            and '$F' in rule.formula[0].upper()
            and rule.dxf is not None
            and rule.dxf.fill is not None
            and rule.dxf.fill.fgColor is not None
            and rule.dxf.fill.fgColor.rgb is not None
            and 'FF0000' in (rule.dxf.fill.fgColor.rgb or '').upper()
        ]
        amber_rules = [
            rule
            for cf_range in cf_rules
            for rule in cf_range.rules
            if rule.type == 'expression'
            and rule.formula
            and 'AMBER' in normalize_formula(rule.formula[0])
            and '$F' in rule.formula[0].upper()
            and rule.dxf is not None
            and rule.dxf.fill is not None
            and rule.dxf.fill.fgColor is not None
            and rule.dxf.fill.fgColor.rgb is not None
            and 'FFFF00' in (rule.dxf.fill.fgColor.rgb or '').upper()
        ]
        green_rules = [
            rule
            for cf_range in cf_rules
            for rule in cf_range.rules
            if rule.type == 'expression'
            and rule.formula
            and 'GREEN' in normalize_formula(rule.formula[0])
            and '$F' in rule.formula[0].upper()
            and rule.dxf is not None
            and rule.dxf.fill is not None
            and rule.dxf.fill.fgColor is not None
            and rule.dxf.fill.fgColor.rgb is not None
            and '00FF00' in (rule.dxf.fill.fgColor.rgb or '').upper()
        ]

        # Also verify at least one CF range covers A2:F13
        cf_covers_range = any(
            'A2' in str(cf_range) for cf_range in cf_rules
        )

        if cf_covers_range and len(red_rules) >= 1 and len(amber_rules) >= 1 and len(green_rules) >= 1:
            print(f"PASS: Component 4 — Conditional formatting on A2:F13 with Red/Amber/Green fills (0.15 pts)")
            total_score += 0.15
        else:
            print(
                f"FAIL: Component 4 — CF covers A2={cf_covers_range}, "
                f"Red rules={len(red_rules)}, Amber rules={len(amber_rules)}, Green rules={len(green_rules)}"
            )
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -----------------------------------------------------------------------
    # Component 5: Freeze row 1 — freeze_panes == "A2" (0.10 points)
    # -----------------------------------------------------------------------
    try:
        fp = ws.freeze_panes
        if fp is not None and str(fp) == 'A2':
            print(f"PASS: Component 5 — freeze_panes='A2' (row 1 frozen) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 — Expected freeze_panes='A2', found: {repr(fp)}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
