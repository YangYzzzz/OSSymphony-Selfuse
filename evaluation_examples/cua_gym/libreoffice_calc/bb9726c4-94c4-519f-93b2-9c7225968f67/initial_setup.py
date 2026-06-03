"""
Initial Setup: Project tracking milestone spreadsheet
Task ID: calc_ops_project_tracking_milestone_013
Domain: libreoffice_calc

Creates a Milestones sheet with 12 project milestones.
Columns A-C are filled; D (Status), E (Days Remaining), F (RAG Status) are empty.
No dropdown, no formulas, no conditional formatting, no freeze pane.
"""

import os
import openpyxl
from datetime import date, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_project_tracking_milestone_013'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Milestones'

    # Headers
    headers = ['Milestone', 'Owner', 'Due Date', 'Status', 'Days Remaining', 'RAG Status']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Base date for relative due dates (today is 2026-03-04)
    today = date(2026, 3, 4)

    # 12 realistic project milestones for a product launch project
    milestones = [
        ('Market Research Complete',        'Sarah Chen',       today - timedelta(days=10)),
        ('Product Requirements Document',   'Marcus Johnson',   today - timedelta(days=3)),
        ('UI/UX Design Sign-off',           'Priya Patel',      today + timedelta(days=2)),
        ('Backend Architecture Review',     'David Kim',        today + timedelta(days=5)),
        ('Alpha Prototype Ready',           'Elena Rodriguez',  today + timedelta(days=14)),
        ('Internal QA Testing',             'James Okonkwo',    today + timedelta(days=21)),
        ('Security Audit Passed',           'Aisha Thompson',   today + timedelta(days=28)),
        ('Beta Release to Partners',        'Marcus Johnson',   today + timedelta(days=35)),
        ('Performance Benchmarking',        'David Kim',        today + timedelta(days=42)),
        ('Documentation Finalized',         'Sarah Chen',       today + timedelta(days=49)),
        ('Regulatory Compliance Check',     'Aisha Thompson',   today + timedelta(days=56)),
        ('Public Launch',                   'Elena Rodriguez',  today + timedelta(days=90)),
    ]

    for r, (milestone, owner, due_date) in enumerate(milestones, 2):
        ws.cell(row=r, column=1, value=milestone)
        ws.cell(row=r, column=2, value=owner)
        ws.cell(row=r, column=3, value=due_date)
        ws.cell(row=r, column=3).number_format = 'yyyy-mm-dd'
        # Columns D, E, F intentionally left empty

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Columns A-C filled for 12 milestones. D (Status), E (Days Remaining), F (RAG Status) are empty.')


create_initial()
