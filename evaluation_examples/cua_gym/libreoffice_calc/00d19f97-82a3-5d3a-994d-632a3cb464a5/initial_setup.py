"""
Initial Setup: SUMPRODUCT with three criteria for hours worked
Task ID: calc_lf_028
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_028'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: TimeLog ---
    ws = wb.active
    ws.title = 'TimeLog'

    # Headers
    ws['A1'] = 'Employee'
    ws['B1'] = 'Dept'
    ws['C1'] = 'Project'
    ws['D1'] = 'Hours'

    # Data rows
    data = [
        ['Chen',  'Engineering', 'Alpha', 40],
        ['Chen',  'Engineering', 'Beta',  25],
        ['Park',  'Engineering', 'Alpha', 35],
        ['Chen',  'Engineering', 'Alpha', 30],
        ['Chen',  'Marketing',   'Alpha', 15],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Result label — F2 intentionally left empty (task asks agent to enter formula)
    ws['F1'] = 'Result'

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['F'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
