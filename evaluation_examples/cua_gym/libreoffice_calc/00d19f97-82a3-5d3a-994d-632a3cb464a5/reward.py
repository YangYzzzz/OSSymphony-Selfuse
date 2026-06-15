"""
Reward Script: SUMPRODUCT with three criteria for total hours
Task ID: calc_lf_028
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): F2 contains a SUMPRODUCT formula
  Component 2 (0.3): Formula includes all three criteria (Chen, Engineering, Alpha)
  Component 3 (0.3): Formula references correct data ranges and Hours column
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_028'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: TimeLog sheet must exist
    if 'TimeLog' not in wb.sheetnames:
        print("FAIL: 'TimeLog' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['TimeLog']

    # Get F2 value (should be a formula string)
    f2_value = ws['F2'].value

    if f2_value is None:
        print(f"FAIL: F2 is empty (no formula entered)")
        print("REWARD: 0.0")
        return 0.0

    f2_str = str(f2_value).strip()
    # Normalize for comparison: uppercase, remove spaces
    f2_norm = f2_str.upper().replace(" ", "")

    # Component 1: F2 contains a SUMPRODUCT formula (0.4 points)
    try:
        if f2_norm.startswith("=SUMPRODUCT("):
            print(f"PASS: Component 1 -- F2 contains SUMPRODUCT formula (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 -- Expected SUMPRODUCT formula in F2, found: {f2_str}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Formula contains all three criteria (Chen, Engineering, Alpha) (0.3 points)
    try:
        has_chen = '"CHEN"' in f2_norm or "'CHEN'" in f2_norm
        has_engineering = '"ENGINEERING"' in f2_norm or "'ENGINEERING'" in f2_norm
        has_alpha = '"ALPHA"' in f2_norm or "'ALPHA'" in f2_norm

        if has_chen and has_engineering and has_alpha:
            print(f"PASS: Component 2 -- All three criteria found (Chen, Engineering, Alpha) (0.3 pts)")
            total_score += 0.3
        else:
            missing = []
            if not has_chen:
                missing.append("Chen")
            if not has_engineering:
                missing.append("Engineering")
            if not has_alpha:
                missing.append("Alpha")
            print(f"FAIL: Component 2 -- Missing criteria in formula: {', '.join(missing)}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Formula references correct ranges including Hours column D (0.3 points)
    # The formula should multiply criteria arrays against D column (Hours)
    # Expected pattern: references to column A (Employee), B (Dept), C (Project), D (Hours)
    try:
        # Check that the formula references ranges in columns A, B, C, and D
        has_a_range = bool(re.search(r'A\d+:A\d+', f2_norm))
        has_b_range = bool(re.search(r'B\d+:B\d+', f2_norm))
        has_c_range = bool(re.search(r'C\d+:C\d+', f2_norm))
        has_d_range = bool(re.search(r'D\d+:D\d+', f2_norm))

        if has_a_range and has_b_range and has_c_range and has_d_range:
            print(f"PASS: Component 3 -- Formula references A, B, C, D column ranges (0.3 pts)")
            total_score += 0.3
        else:
            missing_cols = []
            if not has_a_range:
                missing_cols.append("A (Employee)")
            if not has_b_range:
                missing_cols.append("B (Dept)")
            if not has_c_range:
                missing_cols.append("C (Project)")
            if not has_d_range:
                missing_cols.append("D (Hours)")
            print(f"FAIL: Component 3 -- Missing column ranges: {', '.join(missing_cols)}")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
