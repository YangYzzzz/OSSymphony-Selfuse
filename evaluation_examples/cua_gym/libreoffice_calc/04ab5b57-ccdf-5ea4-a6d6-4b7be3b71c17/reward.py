"""
Reward Script: Build a sales leaderboard with conditional formatting
Task ID: calc_sales_026
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): Attainment formulas in E2:E6 (=Dn/Cn pattern)
  Component 2 (0.25): Percentage number format on E2:E6
  Component 3 (0.40): Conditional formatting rules on E2:E6 (green>=1, red<1)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_026'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get the Leaderboard sheet
    if 'Leaderboard' not in wb.sheetnames:
        print("FAIL: 'Leaderboard' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Leaderboard']

    # Component 1: Attainment formulas in E2:E6 (0.35 points)
    # Each cell should contain a formula like =D{row}/C{row}
    # This is a task-introduced change: initial_env has no values in E column
    try:
        formula_count = 0
        expected_formulas = {
            2: '=D2/C2',
            3: '=D3/C3',
            4: '=D4/C4',
            5: '=D5/C5',
            6: '=D6/C6',
        }
        for row, expected in expected_formulas.items():
            cell = ws.cell(row=row, column=5)
            val = cell.value
            if val is not None and isinstance(val, str):
                # Normalize: remove spaces, uppercase
                normalized = val.upper().replace(" ", "")
                expected_norm = expected.upper().replace(" ", "")
                if normalized == expected_norm:
                    formula_count += 1
                    print(f"  PASS: E{row} has correct formula: {val}")
                else:
                    print(f"  FAIL: E{row} has formula '{val}', expected '{expected}'")
            elif val is not None and isinstance(val, (int, float)):
                # The formula may have been evaluated — check if value is correct ratio
                # D/C ratios: 275000/250000=1.1, 210000/200000=1.05, 185000/200000=0.925,
                #             120000/150000=0.8, 140000/175000=0.8
                expected_ratios = {2: 1.1, 3: 1.05, 4: 0.925, 5: 0.8, 6: 0.8}
                if row in expected_ratios and abs(float(val) - expected_ratios[row]) < 0.01:
                    formula_count += 1
                    print(f"  PASS: E{row} has correct computed value: {val}")
                else:
                    print(f"  FAIL: E{row} has value {val}, expected ~{expected_ratios.get(row)}")
            else:
                print(f"  FAIL: E{row} is empty or unexpected type: {repr(val)}")

        if formula_count == 5:
            print(f"PASS: Component 1 — All 5 attainment formulas correct (0.35 pts)")
            total_score += 0.35
        elif formula_count >= 3:
            partial = round(0.35 * formula_count / 5, 2)
            print(f"PARTIAL: Component 1 — {formula_count}/5 formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {formula_count}/5 formulas correct (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Percentage number format on E2:E6 (0.25 points)
    # Task-introduced change: initial_env has General format (no E values at all)
    try:
        pct_count = 0
        for row in range(2, 7):
            cell = ws.cell(row=row, column=5)
            nf = cell.number_format
            # Check for any percentage format
            if nf and '%' in str(nf):
                pct_count += 1
                print(f"  PASS: E{row} number_format is percentage: '{nf}'")
            else:
                print(f"  FAIL: E{row} number_format is '{nf}', expected percentage")

        if pct_count == 5:
            print(f"PASS: Component 2 — All 5 cells have percentage format (0.25 pts)")
            total_score += 0.25
        elif pct_count >= 3:
            partial = round(0.25 * pct_count / 5, 2)
            print(f"PARTIAL: Component 2 — {pct_count}/5 cells have percentage format ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {pct_count}/5 cells have percentage format (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Conditional formatting rules on E2:E6 (0.40 points)
    # Task-introduced change: initial_env has NO conditional formatting
    # Golden should have: green fill for >= 1 (100%), red fill for < 1
    try:
        green_rule_count = 0
        red_rule_count = 0
        cf_on_e_count = 0

        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            # Check if the conditional formatting applies to E2:E6 range
            if 'E' in cf_range:
                cf_on_e_count += 1
                for rule in cf.rules:
                    if rule.type == 'cellIs' and rule.dxf and rule.dxf.fill:
                        fg_color = None
                        try:
                            fg_color = rule.dxf.fill.fgColor.rgb if rule.dxf.fill.fgColor else None
                        except Exception:
                            pass

                        op = rule.operator
                        formula = rule.formula

                        # Check for green rule (>= 1)
                        if op in ('greaterThanOrEqual', 'greaterThan'):
                            if formula and any('1' in str(f) for f in formula):
                                if fg_color and fg_color.upper() != '00000000':
                                    # Verify greenish color (G channel dominant)
                                    r_val = int(fg_color[2:4], 16)
                                    g_val = int(fg_color[4:6], 16)
                                    b_val = int(fg_color[6:8], 16)
                                    if g_val > r_val and g_val > b_val:
                                        green_rule_count += 1
                                        print(f"  PASS: Green conditional format rule found (operator={op}, color={fg_color})")
                                    else:
                                        print(f"  INFO: Rule with operator={op} has color {fg_color} (not green)")

                        # Check for red rule (< 1)
                        if op in ('lessThan', 'lessThanOrEqual'):
                            if formula and any('1' in str(f) for f in formula):
                                if fg_color and fg_color.upper() != '00000000':
                                    # Verify reddish color (R channel dominant)
                                    r_val = int(fg_color[2:4], 16)
                                    g_val = int(fg_color[4:6], 16)
                                    b_val = int(fg_color[6:8], 16)
                                    if r_val > g_val and r_val > b_val:
                                        red_rule_count += 1
                                        print(f"  PASS: Red conditional format rule found (operator={op}, color={fg_color})")
                                    else:
                                        print(f"  INFO: Rule with operator={op} has color {fg_color} (not red)")

        if green_rule_count >= 1 and red_rule_count >= 1:
            print(f"PASS: Component 3 — Both green (>=1) and red (<1) conditional formatting rules found (0.40 pts)")
            total_score += 0.40
        elif green_rule_count >= 1 or red_rule_count >= 1:
            print(f"PARTIAL: Component 3 — Only one conditional formatting rule found (0.20 pts)")
            total_score += 0.20
        elif cf_on_e_count >= 1:
            print(f"PARTIAL: Component 3 — Conditional formatting exists on E column but rules don't match expected pattern (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — No conditional formatting found on E column (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
