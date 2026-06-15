"""
Initial Setup: Build a sales leaderboard spreadsheet (pre-task state)
Task ID: calc_sales_026
Domain: libreoffice_calc

Initial state has the Leaderboard sheet with raw data but NO attainment
formulas and NO conditional formatting — the agent must add those.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_026'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Leaderboard ---
    ws = wb.active
    ws.title = 'Leaderboard'

    # Headers
    headers = ['Rank', 'Sales Rep', 'Quota', 'Actual', 'Attainment']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    white_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows — raw data only, NO formulas in column E
    data = [
        [1, 'Carol',  250000, 275000],
        [2, 'Eve',    200000, 210000],
        [3, 'Alice',  200000, 185000],
        [4, 'Bob',    150000, 120000],
        [5, 'Dan',    175000, 140000],
    ]

    currency_fmt = '$#,##0'
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(horizontal="center")
            if c in (3, 4):  # Quota and Actual columns
                cell.number_format = currency_fmt

    # Column E (Attainment) is intentionally left EMPTY — the task is to fill it

    # Column widths for readability
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
