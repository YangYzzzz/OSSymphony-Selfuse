"""
Initial Setup: Add a Data Bar conditional formatting rule to Sales Amount column
Task ID: calc_gg3_018
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_018'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sales Sheet ---
    ws = wb.active
    ws.title = 'Sales'

    # Headers
    headers = ['Region', 'Product', 'Sales Amount']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 25 rows of realistic sales data (rows 2-26)
    # Sales amounts range from $1,200 to $98,700
    data = [
        ['Northeast', 'Enterprise Server License', 98700],
        ['West Coast', 'Cloud Storage Plan', 87450],
        ['Southeast', 'Network Security Suite', 76200],
        ['Midwest', 'Database Management Tool', 68900],
        ['Southwest', 'Project Management SaaS', 62300],
        ['Northeast', 'API Gateway Service', 55800],
        ['Pacific Northwest', 'Analytics Dashboard', 51200],
        ['Mid-Atlantic', 'CRM Platform', 47600],
        ['Great Lakes', 'DevOps Automation', 43900],
        ['Southern', 'Video Conferencing', 39500],
        ['Mountain West', 'Email Marketing Suite', 35700],
        ['Central', 'Inventory Management', 32100],
        ['Northeast', 'Backup & Recovery', 28400],
        ['West Coast', 'HR Management System', 25600],
        ['Southeast', 'E-commerce Platform', 22800],
        ['Midwest', 'Accounting Software', 19500],
        ['Southwest', 'Customer Support Tool', 17200],
        ['Pacific Northwest', 'Content Management', 14800],
        ['Mid-Atlantic', 'Payroll Processing', 12500],
        ['Great Lakes', 'Document Signing', 9800],
        ['Southern', 'Fleet Management', 7600],
        ['Mountain West', 'Compliance Tracker', 5400],
        ['Central', 'Appointment Scheduler', 3200],
        ['Northeast', 'Survey Builder', 1800],
        ['West Coast', 'Badge Printing Software', 1200],
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        ws.cell(row=r, column=2, value=row_data[1])
        ws.cell(row=r, column=3, value=row_data[2])

    # Format Sales Amount column as currency
    for r in range(2, 27):
        ws.cell(row=r, column=3).number_format = '$#,##0'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18

    # Bold headers
    from openpyxl.styles import Font
    for col in range(1, 4):
        ws.cell(row=1, column=col).font = Font(bold=True)

    # NO conditional formatting, NO data bars - that's the task

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
