"""
Initial Setup: Multi-currency travel budget planner
Task ID: calc_wf_050
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_050'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Styling helpers ---
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    def style_header(ws, row, max_col):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    # ===== Sheet 1: Rates =====
    ws_rates = wb.active
    ws_rates.title = "Rates"

    rates_headers = ["Currency Code", "Currency Name", "Rate to USD"]
    for col, h in enumerate(rates_headers, 1):
        ws_rates.cell(row=1, column=col, value=h)
    style_header(ws_rates, 1, len(rates_headers))

    currencies = [
        ["EUR", "Euro", 1.0856],
        ["GBP", "British Pound", 1.2634],
        ["JPY", "Japanese Yen", 0.006712],
        ["THB", "Thai Baht", 0.02817],
        ["MXN", "Mexican Peso", 0.05842],
        ["CAD", "Canadian Dollar", 0.7389],
        ["AUD", "Australian Dollar", 0.6523],
        ["CHF", "Swiss Franc", 1.1247],
    ]
    for r, row_data in enumerate(currencies, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_rates.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 3:
                cell.number_format = '0.000000'

    ws_rates.column_dimensions["A"].width = 16
    ws_rates.column_dimensions["B"].width = 22
    ws_rates.column_dimensions["C"].width = 16

    # ===== Sheet 2: Budget =====
    ws_budget = wb.create_sheet("Budget")

    budget_headers = [
        "Date", "Category", "Description", "Local Currency",
        "Amount (Local)", "Rate", "Amount (USD)"
    ]
    for col, h in enumerate(budget_headers, 1):
        ws_budget.cell(row=1, column=col, value=h)
    style_header(ws_budget, 1, len(budget_headers))

    # 10-day trip expenses (multiple per day, realistic travel data)
    expenses = [
        # Day 1 - Tokyo
        ["2025-09-01", "Accommodation", "Hotel Sakura Tokyo - 1 night", "JPY", 18500],
        ["2025-09-01", "Food", "Ramen at Ichiran Shibuya", "JPY", 1290],
        ["2025-09-01", "Transport", "Airport Limousine Bus to Shinjuku", "JPY", 3200],
        # Day 2 - Tokyo
        ["2025-09-02", "Food", "Breakfast at hotel buffet", "JPY", 2800],
        ["2025-09-02", "Activities", "Teamlab Borderless admission", "JPY", 3800],
        ["2025-09-02", "Food", "Sushi omakase at Tsukiji", "JPY", 5500],
        ["2025-09-02", "Transport", "Tokyo Metro day pass", "JPY", 600],
        # Day 3 - Tokyo to Bangkok
        ["2025-09-03", "Transport", "Narita Express to airport", "JPY", 3250],
        ["2025-09-03", "Food", "Airport bento box lunch", "JPY", 1100],
        ["2025-09-03", "Accommodation", "Riverside Hotel Bangkok", "THB", 3200],
        # Day 4 - Bangkok
        ["2025-09-04", "Food", "Street food at Yaowarat Rd", "THB", 350],
        ["2025-09-04", "Activities", "Grand Palace entrance fee", "THB", 500],
        ["2025-09-04", "Transport", "Tuk-tuk rides around old city", "THB", 800],
        ["2025-09-04", "Shopping", "Night market souvenirs", "THB", 1500],
        # Day 5 - Bangkok
        ["2025-09-05", "Food", "Cooking class with lunch", "THB", 2500],
        ["2025-09-05", "Activities", "Floating market day trip", "THB", 1800],
        ["2025-09-05", "Food", "Dinner at Gaggan Anand", "THB", 6500],
        # Day 6 - London
        ["2025-09-06", "Transport", "Heathrow Express to Paddington", "GBP", 25.00],
        ["2025-09-06", "Accommodation", "The Kensington Hotel - 2 nights", "GBP", 189.00],
        ["2025-09-06", "Food", "Fish and chips at Poppies", "GBP", 16.50],
        # Day 7 - London
        ["2025-09-07", "Activities", "Tower of London tickets", "GBP", 33.60],
        ["2025-09-07", "Food", "Afternoon tea at The Ritz", "GBP", 75.00],
        ["2025-09-07", "Transport", "Oyster card top-up", "GBP", 20.00],
        ["2025-09-07", "Shopping", "Books at Daunt Books", "GBP", 28.50],
        # Day 8 - Paris
        ["2025-09-08", "Transport", "Eurostar London to Paris", "EUR", 89.00],
        ["2025-09-08", "Accommodation", "Hotel Le Marais Paris", "EUR", 165.00],
        ["2025-09-08", "Food", "Dinner at Le Bouillon Chartier", "EUR", 32.00],
        # Day 9 - Paris
        ["2025-09-09", "Activities", "Louvre Museum entrance", "EUR", 22.00],
        ["2025-09-09", "Food", "Crepes near Montmartre", "EUR", 8.50],
        ["2025-09-09", "Shopping", "Macarons at Laduree", "EUR", 24.00],
        ["2025-09-09", "Transport", "Metro carnet tickets", "EUR", 16.90],
        # Day 10 - Paris to home
        ["2025-09-10", "Food", "Croissants and coffee at cafe", "EUR", 12.50],
        ["2025-09-10", "Transport", "Taxi to CDG airport", "EUR", 55.00],
        ["2025-09-10", "Shopping", "Duty free perfume", "EUR", 78.00],
    ]

    for r, row_data in enumerate(expenses, 2):
        ws_budget.cell(row=r, column=1, value=row_data[0])
        ws_budget.cell(row=r, column=1).number_format = 'yyyy-mm-dd'
        ws_budget.cell(row=r, column=2, value=row_data[1])
        ws_budget.cell(row=r, column=3, value=row_data[2])
        ws_budget.cell(row=r, column=4, value=row_data[3])
        ws_budget.cell(row=r, column=5, value=row_data[4])
        ws_budget.cell(row=r, column=5).number_format = '#,##0.00'
        # Column 6 (Rate) - LEFT EMPTY (task: add VLOOKUP)
        # Column 7 (Amount USD) - LEFT EMPTY (task: add conversion formula)
        for c in range(1, 8):
            ws_budget.cell(row=r, column=c).border = thin_border

    ws_budget.column_dimensions["A"].width = 14
    ws_budget.column_dimensions["B"].width = 16
    ws_budget.column_dimensions["C"].width = 38
    ws_budget.column_dimensions["D"].width = 16
    ws_budget.column_dimensions["E"].width = 16
    ws_budget.column_dimensions["F"].width = 14
    ws_budget.column_dimensions["G"].width = 16

    # ===== Sheet 3: Summary (empty template - task requires building it) =====
    ws_summary = wb.create_sheet("Summary")

    summary_headers = ["Date", "Daily Total (USD)", "Planned Budget", "Over/Under", "Running Total"]
    for col, h in enumerate(summary_headers, 1):
        ws_summary.cell(row=1, column=col, value=h)
    style_header(ws_summary, 1, len(summary_headers))

    # Add dates for the 10 days (but NO formulas or totals - that's the task)
    for i in range(10):
        day_str = f"2025-09-{1+i:02d}"
        cell = ws_summary.cell(row=2 + i, column=1, value=day_str)
        cell.number_format = 'yyyy-mm-dd'
        cell.border = thin_border
        # Planned budget column pre-filled with $200
        budget_cell = ws_summary.cell(row=2 + i, column=3, value=200)
        budget_cell.number_format = '$#,##0.00'
        budget_cell.border = thin_border
        # Other columns left empty for the task
        for c in [2, 4, 5]:
            ws_summary.cell(row=2 + i, column=c).border = thin_border

    # Category summary section header (row 14)
    ws_summary.cell(row=14, column=1, value="Category Summary")
    ws_summary.cell(row=14, column=1).font = Font(bold=True, size=12)

    cat_headers = ["Category", "Total (USD)"]
    for col, h in enumerate(cat_headers, 1):
        cell = ws_summary.cell(row=15, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    categories = ["Accommodation", "Food", "Transport", "Activities", "Shopping"]
    for r, cat in enumerate(categories, 16):
        cell = ws_summary.cell(row=r, column=1, value=cat)
        cell.border = thin_border
        ws_summary.cell(row=r, column=2).border = thin_border
        # Total column left empty - task requires SUMIF formulas

    ws_summary.column_dimensions["A"].width = 18
    ws_summary.column_dimensions["B"].width = 20
    ws_summary.column_dimensions["C"].width = 18
    ws_summary.column_dimensions["D"].width = 16
    ws_summary.column_dimensions["E"].width = 18

    # Freeze header row on Budget sheet
    ws_budget.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
