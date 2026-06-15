"""
Reward Script: Multi-currency travel budget planner
Task ID: calc_wf_050
Domain: libreoffice_calc
Scoring:
  Component 1: VLOOKUP formulas in Budget Rate column F (0.25)
  Component 2: USD Amount formulas in Budget column G (0.20)
  Component 3: Daily total formulas in Summary B2:B11 (0.15)
  Component 4: Over/Under and Running Total formulas in Summary D,E (0.10)
  Component 5: Category SUMIF totals and Grand Total in Summary (0.10)
  Component 6: Chart present on Summary sheet (0.10)
  Component 7: Conditional formatting on daily totals >200 (0.10)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_050'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice documents."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            import time
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: required sheets exist
    if 'Budget' not in wb.sheetnames or 'Rates' not in wb.sheetnames:
        print("FAIL: Required sheets 'Budget' and/or 'Rates' missing")
        print("REWARD: 0.0")
        return 0.0

    ws_budget = wb['Budget']
    ws_rates = wb['Rates']

    # Find Summary sheet (may not exist in initial)
    has_summary = 'Summary' in wb.sheetnames
    ws_summary = wb['Summary'] if has_summary else None

    # Component 1: VLOOKUP formulas in Budget Rate column F (0.25 points)
    # In initial, F2:F35 are all None. In golden, they contain VLOOKUP formulas.
    try:
        vlookup_count = 0
        total_f_cells = 0
        for row_num in range(2, ws_budget.max_row + 1):
            cell_val = ws_budget.cell(row=row_num, column=6).value  # column F
            d_val = ws_budget.cell(row=row_num, column=4).value  # column D (currency)
            if d_val is not None:  # only count rows that have data
                total_f_cells += 1
                if cell_val is not None and isinstance(cell_val, str) and 'VLOOKUP' in cell_val.upper():
                    vlookup_count += 1

        if total_f_cells > 0 and vlookup_count >= total_f_cells * 0.8:
            print(f"PASS: Component 1 - VLOOKUP in Rate column: {vlookup_count}/{total_f_cells} rows (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 - VLOOKUP in Rate column: {vlookup_count}/{total_f_cells} rows")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: USD Amount formulas in Budget column G (0.20 points)
    # In initial, G2:G35 are all None. In golden, they contain =E*F multiplication formulas.
    try:
        usd_formula_count = 0
        total_g_cells = 0
        for row_num in range(2, ws_budget.max_row + 1):
            cell_val = ws_budget.cell(row=row_num, column=7).value  # column G
            d_val = ws_budget.cell(row=row_num, column=4).value  # column D (currency)
            if d_val is not None:
                total_g_cells += 1
                if cell_val is not None and isinstance(cell_val, str) and '=' in str(cell_val):
                    # Check it references E and F columns (multiplication formula)
                    val_upper = cell_val.upper()
                    if ('E' in val_upper and 'F' in val_upper) or '*' in val_upper:
                        usd_formula_count += 1

        if total_g_cells > 0 and usd_formula_count >= total_g_cells * 0.8:
            print(f"PASS: Component 2 - USD amount formulas in G: {usd_formula_count}/{total_g_cells} rows (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 - USD amount formulas in G: {usd_formula_count}/{total_g_cells} rows")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Daily total formulas in Summary B2:B11 (0.15 points)
    # In initial, Summary B2:B11 are all None. In golden, they have SUMPRODUCT or SUMIFS formulas.
    try:
        daily_total_count = 0
        if ws_summary is not None:
            for row_num in range(2, 12):  # B2:B11
                cell_val = ws_summary.cell(row=row_num, column=2).value
                if cell_val is not None and isinstance(cell_val, str) and '=' in str(cell_val):
                    val_upper = cell_val.upper()
                    # Accept SUMPRODUCT, SUMIFS, SUMIF, or any aggregation formula
                    if any(fn in val_upper for fn in ['SUMPRODUCT', 'SUMIFS', 'SUMIF', 'SUM']):
                        daily_total_count += 1

        if daily_total_count >= 8:  # at least 8 of 10 days have formulas
            print(f"PASS: Component 3 - Daily total formulas: {daily_total_count}/10 (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 - Daily total formulas: {daily_total_count}/10")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Over/Under (D) and Running Total (E) formulas in Summary (0.10 points)
    # In initial, D2:D11 and E2:E11 are all None.
    try:
        over_under_count = 0
        running_total_count = 0
        if ws_summary is not None:
            for row_num in range(2, 12):
                d_val = ws_summary.cell(row=row_num, column=4).value  # column D
                e_val = ws_summary.cell(row=row_num, column=5).value  # column E
                if d_val is not None and isinstance(d_val, str) and '=' in str(d_val):
                    over_under_count += 1
                if e_val is not None and isinstance(e_val, str) and '=' in str(e_val):
                    running_total_count += 1

        if over_under_count >= 8 and running_total_count >= 8:
            print(f"PASS: Component 4 - Over/Under: {over_under_count}/10, Running Total: {running_total_count}/10 (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 - Over/Under: {over_under_count}/10, Running Total: {running_total_count}/10")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Category SUMIF totals and Grand Total (0.10 points)
    # In initial, B16:B20 and B22 are all None.
    try:
        cat_formula_count = 0
        if ws_summary is not None:
            for row_num in range(16, 21):  # B16:B20 category totals
                cell_val = ws_summary.cell(row=row_num, column=2).value
                if cell_val is not None and isinstance(cell_val, str) and '=' in str(cell_val):
                    val_upper = cell_val.upper()
                    if any(fn in val_upper for fn in ['SUMIF', 'SUMPRODUCT', 'SUM']):
                        cat_formula_count += 1

            # Check for grand total somewhere in rows 21-25
            grand_total_matches = [
                row_num for row_num in range(21, 26)
                if (ws_summary.cell(row=row_num, column=2).value is not None
                    and isinstance(ws_summary.cell(row=row_num, column=2).value, str)
                    and 'SUM' in ws_summary.cell(row=row_num, column=2).value.upper())
            ]

        if cat_formula_count >= 4 and len(grand_total_matches) > 0:
            print(f"PASS: Component 5 - Category SUMIF: {cat_formula_count}/5, Grand Total: found (0.10 pts)")
            total_score += 0.10
        elif cat_formula_count >= 4:
            print(f"PARTIAL: Component 5 - Category SUMIF: {cat_formula_count}/5, Grand Total: missing (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 - Category SUMIF: {cat_formula_count}/5, Grand Total: {'found' if len(grand_total_matches) > 0 else 'missing'}")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    # Component 6: Chart on Summary sheet (0.10 points)
    # In initial, there are 0 charts. In golden, there is a chart comparing actual vs planned.
    try:
        chart_count = 0
        if ws_summary is not None:
            chart_count = len(ws_summary._charts)

        if chart_count >= 1:
            print(f"PASS: Component 6 - Chart present on Summary: {chart_count} chart(s) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 6 - No charts on Summary sheet")
    except Exception as e:
        print(f"ERROR: Component 6 - {e}")

    # Component 7: Conditional formatting on daily totals >200 (0.10 points)
    # In initial, there is no conditional formatting. In golden, B2:B11 has cellIs > 200.
    try:
        cf_match_count = 0
        if ws_summary is not None:
            cf_rules = list(ws_summary.conditional_formatting)
            for cf in cf_rules:
                for rule in cf.rules:
                    # Check for a "greater than" type rule with threshold around 200
                    if hasattr(rule, 'operator') and rule.operator in ('greaterThan', 'greaterThanOrEqual'):
                        if rule.formula and len(rule.formula) > 0:
                            try:
                                threshold = float(rule.formula[0])
                                if 195 <= threshold <= 205:
                                    cf_match_count += 1
                            except (ValueError, TypeError):
                                pass
                    # Also accept formula-based rules that reference 200
                    if hasattr(rule, 'type') and rule.type == 'cellIs':
                        if rule.formula and len(rule.formula) > 0:
                            if '200' in str(rule.formula[0]):
                                cf_match_count += 1

        if cf_match_count > 0:
            print(f"PASS: Component 7 - Conditional formatting for >200 found (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 7 - No conditional formatting for daily spending >200")
    except Exception as e:
        print(f"ERROR: Component 7 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
