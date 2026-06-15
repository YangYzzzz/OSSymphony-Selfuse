"""
Reward Script: CSV Import with 'From Row' set to 3
Task ID: calc_gsi_077
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): First row contains actual column headers, not metadata
  Component 2 (0.30): No metadata rows present in the spreadsheet
  Component 3 (0.25): Correct number of data rows (15 employees)
  Component 4 (0.15): Specific data values match expected employee records
"""

import os

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_077'

# Expected column headers (row 3 of original CSV)
EXPECTED_HEADERS = ['Employee ID', 'Full Name', 'Department', 'Hire Date', 'Annual Salary', 'Performance Score']

# Metadata strings that should NOT appear if import skipped first 2 rows
METADATA_MARKERS = ['Export Timestamp', 'System Version', 'AutoReport', 'Environment: PROD']

# Expected first data record (row 4 of original CSV = row 2 of imported xlsx)
EXPECTED_FIRST_RECORD = {
    'A': 'EMP-1001',
    'B': 'Sarah Chen',
    'C': 'Engineering',
}

# Expected last data record (row 18 of original CSV = row 16 of imported xlsx)
EXPECTED_LAST_RECORD = {
    'A': 'EMP-1015',
    'B': 'Yuki Tanaka',
    'C': 'Marketing',
}


def verify_task(file_path):
    """
    Verify that CSV was imported with From Row = 3 (skipping 2 metadata rows).
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Component 1: First row contains actual column headers, not metadata (0.30 points)
    # This is the PRIMARY indicator that From Row was set correctly.
    # If metadata was not skipped, row 1 would contain "Export Timestamp..." instead.
    try:
        row1_values = []
        for col in range(1, 7):
            val = ws.cell(row=1, column=col).value
            row1_values.append(str(val).strip() if val is not None else '')

        # Check that header row matches expected column names
        headers_match = 0
        for i, expected in enumerate(EXPECTED_HEADERS):
            if i < len(row1_values) and row1_values[i].lower() == expected.lower():
                headers_match += 1

        if headers_match >= 5:  # At least 5 of 6 headers match
            print(f"PASS: Component 1 — Row 1 contains correct headers: {row1_values} ({headers_match}/6 match) (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 — Row 1 does not contain expected headers. Found: {row1_values}, matched {headers_match}/6")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: No metadata rows present anywhere in the spreadsheet (0.30 points)
    # If From Row was not set to 3, the metadata rows would be imported as data rows.
    try:
        metadata_found = False
        # Check all cells in first 3 rows for metadata markers
        for row_num in range(1, min(ws.max_row + 1, 5)):
            for col_num in range(1, min(ws.max_column + 1, 7)):
                cell_val = ws.cell(row=row_num, column=col_num).value
                if cell_val is not None:
                    cell_str = str(cell_val)
                    for marker in METADATA_MARKERS:
                        if marker.lower() in cell_str.lower():
                            metadata_found = True
                            print(f"  Found metadata marker '{marker}' in cell ({row_num},{col_num}): {cell_str}")
                            break
                if metadata_found:
                    break
            if metadata_found:
                break

        if not metadata_found:
            print(f"PASS: Component 2 — No metadata rows found in spreadsheet (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 2 — Metadata rows still present (From Row not set to 3?)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Correct number of data rows — 15 employee records (0.25 points)
    # With From Row = 3: 1 header row + 15 data rows = 16 total rows
    try:
        max_row = ws.max_row
        data_rows = max_row - 1  # subtract header row

        if data_rows == 15:
            print(f"PASS: Component 3 — Correct data row count: {data_rows} employee records (0.25 pts)")
            total_score += 0.25
        elif 13 <= data_rows <= 17:
            # Partial credit for close but not exact
            partial = 0.10
            print(f"PARTIAL: Component 3 — Data rows = {data_rows}, expected 15 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Expected 15 data rows, found {data_rows} (max_row={max_row})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Specific data values match expected employee records (0.15 points)
    # Verify first and last records to confirm data alignment is correct
    try:
        checks_passed = 0
        total_checks = 6

        # Check first record (row 2)
        for col_letter, expected_val in EXPECTED_FIRST_RECORD.items():
            actual = ws[f'{col_letter}2'].value
            if actual is not None and str(actual).strip() == expected_val:
                checks_passed += 1
            else:
                print(f"  Data mismatch at {col_letter}2: expected '{expected_val}', found '{actual}'")

        # Check last record (row 16)
        for col_letter, expected_val in EXPECTED_LAST_RECORD.items():
            actual = ws[f'{col_letter}16'].value
            if actual is not None and str(actual).strip() == expected_val:
                checks_passed += 1
            else:
                print(f"  Data mismatch at {col_letter}16: expected '{expected_val}', found '{actual}'")

        if checks_passed >= 5:
            print(f"PASS: Component 4 — Data values correct: {checks_passed}/{total_checks} checks passed (0.15 pts)")
            total_score += 0.15
        elif checks_passed >= 3:
            partial = 0.08
            print(f"PARTIAL: Component 4 — {checks_passed}/{total_checks} data checks passed ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {checks_passed}/{total_checks} data checks passed")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
