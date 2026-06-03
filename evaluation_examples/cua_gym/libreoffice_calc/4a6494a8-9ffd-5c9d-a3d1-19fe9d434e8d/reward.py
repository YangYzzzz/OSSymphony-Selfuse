"""
Reward Script: Fill all total row and total column formulas in the multi-year budget matrix,
               and add percentage-of-total formulas in an additional row below the total row.
Task ID: osworld_calc_fill_totals_007
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Total column formulas (G2:G13) filled with SUM formulas — 0.35 points
  Component 2: Total row formulas (B14:G14) filled with SUM formulas — 0.35 points
  Component 3: Percentage-of-total row added (A15 + B15:F15 % formulas) — 0.30 points
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_fill_totals_007'


def is_sum_formula(value):
    """Check if cell value is a SUM formula."""
    if not isinstance(value, str):
        return False
    normalized = value.strip().upper().replace(' ', '')
    return normalized.startswith('=SUM(')


def is_pct_formula(value):
    """
    Check if cell value is a % of total formula.
    Expected pattern: =Xx/$G$14*100 or similar referencing G14 (the grand total cell).
    More broadly, any formula referencing the total row and dividing by G14 (grand total).
    """
    if not isinstance(value, str):
        return False
    normalized = value.strip().upper().replace(' ', '')
    # Must be a formula
    if not normalized.startswith('='):
        return False
    # Must reference $G$14 or G14 (the grand total cell)
    has_g14_ref = bool(re.search(r'\$?G\$?14', normalized))
    # Must involve a division (percentage calculation)
    has_division = '/' in normalized
    return has_g14_ref and has_division


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: the 'Budget' sheet must exist
    if 'Budget' not in wb.sheetnames:
        print(f"CRITICAL: 'Budget' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Budget']

    # -----------------------------------------------------------------------
    # Component 1: Total column formulas (G2:G13) — 0.35 points
    # The golden file has =SUM(Bx:Fx) in G2 through G13 (one per cost center row).
    # In the initial file these cells are empty.
    # We award full component credit if ALL 12 rows have a SUM formula in column G.
    # -----------------------------------------------------------------------
    try:
        data_rows = range(2, 14)  # rows 2..13 inclusive (12 cost center rows)
        total_col_filled = 0
        total_col_expected = len(data_rows)
        total_col_details = []

        for row in data_rows:
            cell = ws.cell(row=row, column=7)  # column G
            val = cell.value
            if is_sum_formula(val):
                total_col_filled += 1
            else:
                total_col_details.append(f"G{row}={repr(val)}")

        if total_col_filled == total_col_expected:
            print(f"PASS: Component 1 — All {total_col_expected} Total column SUM formulas present in G2:G13 (0.35 pts)")
            total_score += 0.35
        elif total_col_filled > 0:
            partial = round(0.35 * total_col_filled / total_col_expected, 4)
            print(f"PARTIAL: Component 1 — {total_col_filled}/{total_col_expected} Total column SUM formulas present ({partial} pts)")
            print(f"  Missing/wrong cells: {total_col_details[:5]}")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No Total column SUM formulas found in G2:G13")
            print(f"  First few cell values: {total_col_details[:5]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Total row formulas (B14:G14) — 0.35 points
    # The golden file has =SUM(Bx:Bx) in B14 through G14 (one per year column plus grand total).
    # In the initial file row 14 has only the label "Total" in A14 and all other cells empty.
    # We award full component credit if ALL 6 cells (B14:G14) have SUM formulas.
    # -----------------------------------------------------------------------
    try:
        total_row_num = 14
        year_cols = range(2, 8)  # columns B=2 through G=7
        total_row_filled = 0
        total_row_expected = len(year_cols)
        total_row_details = []

        for col in year_cols:
            cell = ws.cell(row=total_row_num, column=col)
            val = cell.value
            if is_sum_formula(val):
                total_row_filled += 1
            else:
                col_letter = openpyxl.utils.get_column_letter(col)
                total_row_details.append(f"{col_letter}{total_row_num}={repr(val)}")

        if total_row_filled == total_row_expected:
            print(f"PASS: Component 2 — All {total_row_expected} Total row SUM formulas present in B14:G14 (0.35 pts)")
            total_score += 0.35
        elif total_row_filled > 0:
            partial = round(0.35 * total_row_filled / total_row_expected, 4)
            print(f"PARTIAL: Component 2 — {total_row_filled}/{total_row_expected} Total row SUM formulas present ({partial} pts)")
            print(f"  Missing/wrong cells: {total_row_details}")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No Total row SUM formulas found in B14:G14")
            print(f"  Cell values: {total_row_details}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Percentage-of-total row (row 15) — 0.30 points
    # The golden file adds a "% of Total" row at row 15:
    #   A15 = "% of Total" label
    #   B15:F15 = formulas like =B14/$G$14*100 (each year's share of grand total)
    #   G15 = =SUM(B15:F15)
    # In the initial file row 15 does not exist.
    # We award points if:
    #   - A15 has a label indicating percentage (0.05 pts)
    #   - B15:F15 each have a formula referencing G14 with division (0.20 pts)
    #   - G15 has a SUM formula (0.05 pts)
    # -----------------------------------------------------------------------
    try:
        pct_score = 0.0

        # Sub-check 3a: A15 has a "% of Total" label
        a15_val = ws.cell(row=15, column=1).value
        if a15_val is not None and '%' in str(a15_val):
            print(f"PASS: Component 3a — A15 label present: {repr(a15_val)} (0.05 pts)")
            pct_score += 0.05
        else:
            print(f"FAIL: Component 3a — A15 label missing or no '%' sign: {repr(a15_val)}")

        # Sub-check 3b: B15:F15 have percentage formulas referencing G14
        pct_cols = range(2, 7)  # columns B=2 through F=6
        pct_filled = 0
        pct_expected = len(pct_cols)
        pct_details = []

        for col in pct_cols:
            cell = ws.cell(row=15, column=col)
            val = cell.value
            if is_pct_formula(val):
                pct_filled += 1
            else:
                col_letter = openpyxl.utils.get_column_letter(col)
                pct_details.append(f"{col_letter}15={repr(val)}")

        if pct_filled == pct_expected:
            print(f"PASS: Component 3b — All {pct_expected} % formulas present in B15:F15 (0.20 pts)")
            pct_score += 0.20
        elif pct_filled > 0:
            partial_pct = round(0.20 * pct_filled / pct_expected, 4)
            print(f"PARTIAL: Component 3b — {pct_filled}/{pct_expected} % formulas present ({partial_pct} pts)")
            print(f"  Missing/wrong cells: {pct_details}")
            pct_score += partial_pct
        else:
            print(f"FAIL: Component 3b — No % formulas found in B15:F15")
            print(f"  Cell values: {pct_details}")

        # Sub-check 3c: G15 has a SUM formula summarizing the % row
        g15_val = ws.cell(row=15, column=7).value
        if is_sum_formula(g15_val):
            print(f"PASS: Component 3c — G15 SUM formula present: {repr(g15_val)} (0.05 pts)")
            pct_score += 0.05
        else:
            print(f"FAIL: Component 3c — G15 SUM formula missing: {repr(g15_val)}")

        total_score += pct_score
        print(f"Component 3 subtotal: {pct_score}/0.30")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path on the VM
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
