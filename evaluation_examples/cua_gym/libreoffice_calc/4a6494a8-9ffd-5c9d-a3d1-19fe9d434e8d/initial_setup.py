"""
Initial Setup: Multi-year budget matrix with empty total row/column
Task ID: osworld_calc_fill_totals_007
Domain: libreoffice_calc

Creates a spreadsheet with:
- Budget matrix: cost centers (rows) x years (columns B-F for 2021-2025)
- Column G header "Total" but ALL cells in that column are EMPTY
- A "Total" row at the bottom with ALL cells EMPTY
- No "% of Total" row (that must be added by the agent)
"""

import os
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_fill_totals_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Budget ---
    ws = wb.active
    ws.title = "Budget"

    # Styles
    header_font = Font(bold=True, size=11, name="Calibri")
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, name="Calibri", color="FFFFFFFF")
    thin = Side(style="thin", color="FF000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    # Row 1: Headers
    # A1: "Cost Center", B1-F1: years, G1: "Total"
    years = [2021, 2022, 2023, 2024, 2025]
    headers_row = ["Cost Center"] + years + ["Total"]
    for col, h in enumerate(headers_row, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Cost center data: realistic business departments with multi-year budgets
    cost_centers = [
        ("Engineering",       245000, 268000, 312000, 345000, 389000),
        ("Marketing",         98000,  112000, 125000, 138000, 155000),
        ("Human Resources",   67000,  71000,  78000,  82000,  89000),
        ("Finance",           54000,  57000,  62000,  66000,  70000),
        ("Operations",        183000, 196000, 211000, 228000, 247000),
        ("Research & Dev",    312000, 345000, 389000, 421000, 467000),
        ("Customer Support",  76000,  82000,  88000,  95000,  103000),
        ("Sales",             142000, 158000, 172000, 189000, 207000),
        ("IT Infrastructure", 221000, 238000, 256000, 275000, 297000),
        ("Legal & Compliance",45000,  48000,  52000,  56000,  61000),
        ("Facilities",        89000,  93000,  98000,  104000, 111000),
        ("Product Mgmt",      134000, 148000, 163000, 179000, 196000),
    ]

    # Alternate row fills for readability
    light_fill = PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")

    for row_idx, (dept, *values) in enumerate(cost_centers, 2):
        fill = light_fill if row_idx % 2 == 0 else white_fill

        # Column A: Cost center name
        cell_a = ws.cell(row=row_idx, column=1, value=dept)
        cell_a.font = Font(name="Calibri", size=10)
        cell_a.fill = fill
        cell_a.alignment = Alignment(horizontal="left", vertical="center")
        cell_a.border = border

        # Columns B-F: Year values
        for col_idx, val in enumerate(values, 2):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=10)
            cell.fill = fill
            cell.alignment = right_align
            cell.number_format = '#,##0'
            cell.border = border

        # Column G: Total — INTENTIONALLY LEFT EMPTY (agent must fill this)
        cell_g = ws.cell(row=row_idx, column=7, value=None)
        cell_g.font = Font(name="Calibri", size=10)
        cell_g.fill = fill
        cell_g.border = border

    # Total row (row 14) — ALL cells EMPTY (agent must fill the SUM formulas)
    total_row = len(cost_centers) + 2  # row 14
    total_label_cell = ws.cell(row=total_row, column=1, value="Total")
    total_label_cell.font = Font(bold=True, name="Calibri", size=10)
    total_label_cell.fill = PatternFill(start_color="FFBDD7EE", end_color="FFBDD7EE", fill_type="solid")
    total_label_cell.alignment = Alignment(horizontal="left", vertical="center")
    total_label_cell.border = border

    for col_idx in range(2, 8):  # Columns B-G all empty
        cell = ws.cell(row=total_row, column=col_idx, value=None)
        cell.font = Font(bold=True, name="Calibri", size=10)
        cell.fill = PatternFill(start_color="FFBDD7EE", end_color="FFBDD7EE", fill_type="solid")
        cell.border = border
        cell.alignment = right_align

    # Column widths
    ws.column_dimensions["A"].width = 22
    for col_letter in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_letter].width = 14

    # Row heights
    ws.row_dimensions[1].height = 22
    for r in range(2, total_row + 1):
        ws.row_dimensions[r].height = 18

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
