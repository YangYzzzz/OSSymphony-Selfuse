"""
Initial Setup: Create spreadsheet with Audit, Log1, Log2, Log3 sheets.
Task ID: calc_mcp_059
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_059'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Audit ---
    ws_audit = wb.active
    ws_audit.title = 'Audit'
    ws_audit.cell(row=1, column=1, value='Metric')
    ws_audit.cell(row=1, column=2, value='Value')
    ws_audit.cell(row=2, column=1, value='Total Entries')
    # B2 intentionally left empty - this is where the agent must place the formula

    # --- Sheet: Log1 (8 entries) ---
    ws_log1 = wb.create_sheet('Log1')
    ws_log1.cell(row=1, column=1, value='Entry')
    log1_data = [
        'Server restart initiated by admin',
        'Disk usage threshold exceeded 85%',
        'Backup completed successfully',
        'New user account created: jmartin',
        'SSL certificate renewed for api.internal',
        'Firewall rule updated: allow port 8443',
        'Scheduled maintenance window opened',
        'Database index rebuild completed',
    ]
    for i, entry in enumerate(log1_data, 2):
        ws_log1.cell(row=i, column=1, value=entry)

    # --- Sheet: Log2 (5 entries) ---
    ws_log2 = wb.create_sheet('Log2')
    ws_log2.cell(row=1, column=1, value='Entry')
    log2_data = [
        'Payment gateway timeout for order #40921',
        'API rate limit reached for client XC-7712',
        'Cache invalidation triggered for product catalog',
        'Email delivery failure: bounced address ops@legacy.net',
        'Load balancer health check failed on node-3',
    ]
    for i, entry in enumerate(log2_data, 2):
        ws_log2.cell(row=i, column=1, value=entry)

    # --- Sheet: Log3 (7 entries) ---
    ws_log3 = wb.create_sheet('Log3')
    ws_log3.cell(row=1, column=1, value='Entry')
    log3_data = [
        'Deployment v2.14.3 rolled out to production',
        'Memory usage spike detected on worker-05',
        'Cron job analytics_daily completed in 47s',
        'User session expired: tkaminski (idle 30min)',
        'Webhook delivery confirmed to partner endpoint',
        'DNS record updated for staging.example.com',
        'Audit log export generated for Q1 2025',
    ]
    for i, entry in enumerate(log3_data, 2):
        ws_log3.cell(row=i, column=1, value=entry)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
