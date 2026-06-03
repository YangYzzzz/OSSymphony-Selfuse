"""
Reward Script: Count non-empty cells across multiple log sheets using COUNTA
Task ID: calc_mcp_059
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Audit!B2 contains a formula (not empty, not a plain value)
  Component 2 (0.3): Formula references all three log sheets (Log1, Log2, Log3) with COUNTA
  Component 3 (0.3): Formula correctly excludes header rows (row 1) by starting at A2 or later
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_059'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Audit sheet must exist
    if 'Audit' not in wb.sheetnames:
        print("FAIL: 'Audit' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Audit']
    b2_value = ws['B2'].value

    # Component 1: Audit!B2 contains a formula (0.4 points)
    # In initial_env, B2 is None. In golden_env, B2 has a formula string.
    try:
        if b2_value is not None and isinstance(b2_value, str) and b2_value.strip().startswith('='):
            formula = b2_value.strip().upper()
            print(f"PASS: Component 1 — B2 contains a formula: {b2_value} (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — B2 should contain a formula, found: {repr(b2_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Formula uses COUNTA and references all three log sheets (0.3 points)
    # The formula must mention Log1, Log2, and Log3 with COUNTA function
    try:
        if b2_value is not None and isinstance(b2_value, str):
            formula_upper = b2_value.strip().upper()
            has_counta = 'COUNTA' in formula_upper
            has_log1 = 'LOG1' in formula_upper
            has_log2 = 'LOG2' in formula_upper
            has_log3 = 'LOG3' in formula_upper

            if has_counta and has_log1 and has_log2 and has_log3:
                print(f"PASS: Component 2 — Formula uses COUNTA across Log1, Log2, Log3 (0.3 pts)")
                total_score += 0.3
            else:
                missing = []
                if not has_counta:
                    missing.append('COUNTA function')
                if not has_log1:
                    missing.append('Log1 reference')
                if not has_log2:
                    missing.append('Log2 reference')
                if not has_log3:
                    missing.append('Log3 reference')
                print(f"FAIL: Component 2 — Missing: {', '.join(missing)}. Formula: {b2_value}")
        else:
            print(f"FAIL: Component 2 — No formula in B2 to check")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Formula excludes header rows by starting at row 2+ (0.3 points)
    # Each COUNTA reference should NOT start at A1 (header). Valid patterns:
    # COUNTA(Log1!A2:...) or COUNTA(Log1!A2:A1000) etc.
    # Invalid: COUNTA(Log1!A1:...) which would include the header
    try:
        if b2_value is not None and isinstance(b2_value, str):
            formula_upper = b2_value.strip().upper().replace(' ', '')
            # Find all sheet references like LOG1!A<row> or LOG2!A<row> etc.
            # Pattern: LOG<n>!A<number> at start of range
            refs = re.findall(r'LOG[123]!A(\d+)', formula_upper)
            if len(refs) >= 3:
                # Check that the starting row references are >= 2 (excluding header)
                # In COUNTA(Log1!A2:A1000), we get ['2', '1000'] for Log1
                # We need to check the start-of-range rows
                # Pattern: COUNTA(LOGn!A<start>:A<end>) - start should be >= 2
                range_starts = re.findall(r'COUNTA\(LOG[123]!A(\d+):', formula_upper)
                if len(range_starts) >= 3:
                    all_exclude_header = all(int(r) >= 2 for r in range_starts)
                    if all_exclude_header:
                        print(f"PASS: Component 3 — All COUNTA ranges start at row 2+, excluding headers (0.3 pts)")
                        total_score += 0.3
                    else:
                        print(f"FAIL: Component 3 — Some ranges start at row 1 (includes header): starts={range_starts}")
                else:
                    # Try alternative formula patterns like COUNTA(Log1!A:A)-1
                    # or using a single COUNTA with multiple ranges
                    # Accept if formula has COUNTA and doesn't reference A1 as start
                    start_refs = re.findall(r'LOG[123]!A(\d+)', formula_upper)
                    # Filter to get unique start-of-range values
                    # If no A1 references found, it's likely correct
                    if start_refs and all(int(r) >= 2 for r in start_refs):
                        print(f"PASS: Component 3 — References start at row 2+, headers excluded (0.3 pts)")
                        total_score += 0.3
                    elif '!A1' not in formula_upper.replace('!A1000', '').replace('!A100', '').replace('!A10', ''):
                        # No A1 reference (after excluding A1000, A100, A10)
                        print(f"PASS: Component 3 — No header row (A1) references found (0.3 pts)")
                        total_score += 0.3
                    else:
                        print(f"FAIL: Component 3 — Cannot confirm header exclusion. Formula: {b2_value}")
            else:
                print(f"FAIL: Component 3 — Fewer than 3 sheet references found. Formula: {b2_value}")
        else:
            print(f"FAIL: Component 3 — No formula in B2 to check")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
