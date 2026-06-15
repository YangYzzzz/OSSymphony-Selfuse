"""
Reward Script: Create a simple pivot table showing total units sold per region.
Task ID: calc_pivot_008
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20): PivotTable sheet exists (new sheet beyond RegionalSales)
  Component 2 (0.50): All 4 regions have correct UnitsSold totals
  Component 3 (0.15): Grand Total row equals 2000
  Component 4 (0.15): Proper headers (Region + UnitsSold-related column)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_008'

# Expected region totals from the task context
EXPECTED_REGIONS = {
    'North': 520,
    'South': 480,
    'East': 610,
    'West': 390,
}
EXPECTED_GRAND_TOTAL = 2000


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Identify the pivot table sheet: any sheet OTHER than 'RegionalSales'
    pivot_sheets = [s for s in wb.sheetnames if s != 'RegionalSales']

    # Component 1: A new pivot table sheet exists (0.20 points)
    # This FAILS on initial (only RegionalSales) and PASSES on golden (has PivotTable)
    try:
        if len(pivot_sheets) >= 1:
            pivot_ws = wb[pivot_sheets[0]]
            print(f"PASS: Component 1 — Pivot sheet '{pivot_sheets[0]}' exists (0.20 pts)")
            total_score += 0.20
        else:
            print("FAIL: Component 1 — No pivot table sheet found (only RegionalSales exists)")
            # Without a pivot sheet, no further checks are possible
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Read pivot sheet data into a dict: region_name -> units_sold value
    region_data = {}
    grand_total_val = None
    header_region = None
    header_units = None

    try:
        # Read headers from row 1
        header_region = pivot_ws.cell(row=1, column=1).value
        header_units = pivot_ws.cell(row=1, column=2).value

        # Read data rows
        for r in range(2, pivot_ws.max_row + 1):
            label = pivot_ws.cell(row=r, column=1).value
            value = pivot_ws.cell(row=r, column=2).value
            if label is None:
                continue
            label_str = str(label).strip()
            # Check if this is the grand total row
            if 'total' in label_str.lower() and 'grand' in label_str.lower():
                grand_total_val = value
            elif label_str in EXPECTED_REGIONS:
                try:
                    region_data[label_str] = float(value)
                except (ValueError, TypeError):
                    region_data[label_str] = None
    except Exception as e:
        print(f"ERROR: Reading pivot data — {e}")

    # Component 2: All 4 regions with correct UnitsSold totals (0.50 points, 0.125 each)
    try:
        region_score = 0.0
        for region, expected_val in EXPECTED_REGIONS.items():
            if region in region_data and region_data[region] is not None:
                if abs(region_data[region] - expected_val) < 0.01:
                    print(f"PASS: Component 2 — {region} = {region_data[region]} (expected {expected_val}) (0.125 pts)")
                    region_score += 0.125
                else:
                    print(f"FAIL: Component 2 — {region} = {region_data[region]}, expected {expected_val}")
            else:
                print(f"FAIL: Component 2 — {region} not found in pivot table")
        if region_score > 0:
            total_score += region_score
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Grand Total row equals 2000 (0.15 points)
    try:
        if grand_total_val is not None:
            if abs(float(grand_total_val) - EXPECTED_GRAND_TOTAL) < 0.01:
                print(f"PASS: Component 3 — Grand Total = {grand_total_val} (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 3 — Grand Total = {grand_total_val}, expected {EXPECTED_GRAND_TOTAL}")
        else:
            print("FAIL: Component 3 — No Grand Total row found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Proper headers (0.15 points)
    # Must have a Region-like header and a UnitsSold-like header
    try:
        # Derive booleans directly from string checks (no bare True assignment)
        has_region_header = (header_region is not None and 'region' in str(header_region).strip().lower())
        has_units_header = (header_units is not None and ('unit' in str(header_units).strip().lower() or 'sold' in str(header_units).strip().lower()))

        if has_region_header and has_units_header:
            print(f"PASS: Component 4 — Headers: '{header_region}' and '{header_units}' (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — Headers: '{header_region}' and '{header_units}' (need Region and UnitsSold)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
