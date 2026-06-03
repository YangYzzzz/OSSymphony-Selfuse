"""
Initial Setup: Create a spreadsheet with regional sales data for pivot table task.
Task ID: calc_pivot_008
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_008'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'RegionalSales'

    # --- Headers ---
    headers = ['TransID', 'Date', 'Region', 'Product', 'UnitsSold', 'Revenue']
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    white_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Data generation ---
    # We need exact regional totals: North=520, South=480, East=610, West=390
    # Total = 2000 across 180 transactions
    regions = ['North', 'South', 'East', 'West']
    target_units = {'North': 520, 'South': 480, 'East': 610, 'West': 390}

    # Distribute 180 transactions across regions
    # North: 45, South: 45, East: 45, West: 45 = 180
    region_counts = {'North': 45, 'South': 45, 'East': 45, 'West': 45}

    products = [
        'Laptop Pro', 'Wireless Mouse', 'USB Hub', 'Monitor 27"',
        'Keyboard MX', 'Webcam HD', 'Headset Elite', 'Docking Station',
        'SSD 1TB', 'RAM 16GB', 'Power Bank', 'Cable Kit'
    ]

    # Generate units for each region that sum to the target
    def distribute_units(total, count):
        """Distribute total units across count transactions."""
        base = total // count
        remainder = total % count
        units = [base] * count
        for i in range(remainder):
            units[i] += 1
        # Shuffle to add variation but keep the sum exact
        random.shuffle(units)
        return units

    # Build all rows
    rows = []
    for region in regions:
        count = region_counts[region]
        units_list = distribute_units(target_units[region], count)
        for i, units in enumerate(units_list):
            rows.append({
                'region': region,
                'product': random.choice(products),
                'units': units,
            })

    # Shuffle all rows to mix regions
    random.shuffle(rows)

    # Generate dates spread across 2025
    start_date = datetime(2025, 1, 5)
    end_date = datetime(2025, 12, 20)
    date_range_days = (end_date - start_date).days

    for idx, row_data in enumerate(rows):
        r = idx + 2  # row number (1-indexed, header is row 1)
        trans_id = f'T{idx + 1:03d}'
        date_val = start_date + timedelta(days=random.randint(0, date_range_days))
        units = row_data['units']
        # Price per unit varies by product type
        price_per_unit = random.uniform(25.0, 350.0)
        revenue = round(units * price_per_unit, 2)

        ws.cell(row=r, column=1, value=trans_id)
        ws.cell(row=r, column=2, value=date_val)
        ws.cell(row=r, column=2).number_format = 'yyyy-mm-dd'
        ws.cell(row=r, column=3, value=row_data['region'])
        ws.cell(row=r, column=4, value=row_data['product'])
        ws.cell(row=r, column=5, value=units)
        ws.cell(row=r, column=6, value=revenue)
        ws.cell(row=r, column=6).number_format = '$#,##0.00'

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Verify totals
    region_totals = {'North': 0, 'South': 0, 'East': 0, 'West': 0}
    for r in range(2, 182):
        region = ws.cell(row=r, column=3).value
        units = ws.cell(row=r, column=5).value
        region_totals[region] += units
    print(f'Verification - Region totals: {region_totals}')
    print(f'Verification - Grand total: {sum(region_totals.values())}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
