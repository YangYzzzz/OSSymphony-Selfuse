"""
Reward Script: Apply square-root curve and Pass/Fail formulas to grade spreadsheet
Task ID: calc_edu_curve_grades_004
Domain: libreoffice_calc
Scoring:
  Component 1 (0.50): Column C (Curved Score) contains =ROUND(SQRT(B)*10,1) formula for all 35 students
  Component 2 (0.30): Column D (Pass/Fail) contains =IF(C>=60,"Pass","Fail") formula for all 35 students
  Component 3 (0.20): Column C cells are formatted to 1 decimal place (number_format '0.0' or equivalent)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_edu_curve_grades_004'

# Expected row range: rows 2-36 (35 students)
DATA_START_ROW = 2
DATA_END_ROW = 36
NUM_STUDENTS = 35


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet exists
    if 'Midterm' not in wb.sheetnames:
        print("FAIL: 'Midterm' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Midterm']

    # Component 1: Column C (Curved Score) contains ROUND(SQRT(B)*10,1) formula
    # for all 35 student rows (rows 2–36). (0.50 points)
    # The formula must reference the corresponding B cell and use ROUND + SQRT + *10
    try:
        c_formula_count = 0
        c_formula_errors = []
        for row in range(DATA_START_ROW, DATA_END_ROW + 1):
            cell_val = ws.cell(row=row, column=3).value
            if cell_val is None:
                c_formula_errors.append(f"Row {row}: C column is empty")
                continue
            val_str = str(cell_val).upper().replace(" ", "")
            # Must contain SQRT(Brow), *10 and ROUND with ,1
            expected_b_ref = f"B{row}"
            if (
                "SQRT(" + expected_b_ref.upper() + ")" in val_str and
                "*10" in val_str and
                "ROUND(" in val_str and
                ",1)" in val_str
            ):
                c_formula_count += 1
            else:
                c_formula_errors.append(
                    f"Row {row}: expected ROUND(SQRT({expected_b_ref})*10,1) pattern, got: {cell_val}"
                )

        if c_formula_count == NUM_STUDENTS:
            print(f"PASS: Component 1 — All {NUM_STUDENTS} rows in column C have correct ROUND(SQRT(B)*10,1) formula (0.50 pts)")
            total_score += 0.50
        elif c_formula_count >= NUM_STUDENTS // 2:
            # Partial credit: more than half correct
            partial = round(0.50 * c_formula_count / NUM_STUDENTS, 2)
            print(f"PARTIAL: Component 1 — {c_formula_count}/{NUM_STUDENTS} rows in column C have correct formula ({partial} pts)")
            print(f"  First error: {c_formula_errors[0] if c_formula_errors else 'N/A'}")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {c_formula_count}/{NUM_STUDENTS} rows have correct ROUND(SQRT) formula in column C")
            if c_formula_errors:
                print(f"  First error: {c_formula_errors[0]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Column D (Pass/Fail) contains =IF(C>=60,"Pass","Fail") for all 35 rows (0.30 points)
    try:
        d_formula_count = 0
        d_formula_errors = []
        for row in range(DATA_START_ROW, DATA_END_ROW + 1):
            cell_val = ws.cell(row=row, column=4).value
            if cell_val is None:
                d_formula_errors.append(f"Row {row}: D column is empty")
                continue
            val_str = str(cell_val).upper().replace(" ", "")
            # Must contain IF(Crow>=60,...) with Pass and Fail
            expected_c_ref = f"C{row}"
            if (
                "IF(" + expected_c_ref.upper() + ">=60" in val_str and
                "PASS" in val_str and
                "FAIL" in val_str
            ):
                d_formula_count += 1
            else:
                d_formula_errors.append(
                    f"Row {row}: expected IF(C{row}>=60,\"Pass\",\"Fail\") pattern, got: {cell_val}"
                )

        if d_formula_count == NUM_STUDENTS:
            print(f"PASS: Component 2 — All {NUM_STUDENTS} rows in column D have correct IF(C>=60,\"Pass\",\"Fail\") formula (0.30 pts)")
            total_score += 0.30
        elif d_formula_count >= NUM_STUDENTS // 2:
            partial = round(0.30 * d_formula_count / NUM_STUDENTS, 2)
            print(f"PARTIAL: Component 2 — {d_formula_count}/{NUM_STUDENTS} rows in column D have correct formula ({partial} pts)")
            print(f"  First error: {d_formula_errors[0] if d_formula_errors else 'N/A'}")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {d_formula_count}/{NUM_STUDENTS} rows have IF Pass/Fail formula in column D")
            if d_formula_errors:
                print(f"  First error: {d_formula_errors[0]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Column C cells formatted to 1 decimal place (0.20 points)
    # Accept number_format of '0.0', '#,##0.0', '0.00' is NOT acceptable (must be 1 decimal)
    # Check at least the first few data rows
    try:
        decimal_format_count = 0
        for row in range(DATA_START_ROW, DATA_END_ROW + 1):
            cell = ws.cell(row=row, column=3)
            fmt = cell.number_format if cell.number_format else 'General'
            # Patterns that indicate 1 decimal: '0.0', '#,##0.0', '0.0%' not valid
            # We check that it is a 1-decimal format
            if re.search(r'0\.0(?!0)', fmt):
                decimal_format_count += 1

        if decimal_format_count == NUM_STUDENTS:
            print(f"PASS: Component 3 — All {NUM_STUDENTS} cells in column C formatted to 1 decimal place (0.20 pts)")
            total_score += 0.20
        elif decimal_format_count > 0:
            partial = round(0.20 * decimal_format_count / NUM_STUDENTS, 2)
            print(f"PARTIAL: Component 3 — {decimal_format_count}/{NUM_STUDENTS} cells in column C have 1-decimal format ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No cells in column C have 1-decimal number format. "
                  f"Sample format: '{ws.cell(row=2, column=3).number_format}'")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
