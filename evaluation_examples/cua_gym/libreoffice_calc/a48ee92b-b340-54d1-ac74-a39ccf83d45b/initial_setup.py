"""
Initial Setup: Apply square-root curve to midterm scores
Task ID: calc_edu_curve_grades_004
Domain: libreoffice_calc

Creates a Midterm sheet with 35 students' names and raw scores (20-95).
Columns C (Curved Score) and D (Pass/Fail) are intentionally empty.
"""

import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_curve_grades_004'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Midterm ---
    ws = wb.active
    ws.title = 'Midterm'

    # Headers
    headers = ['Student Name', 'Raw Score', 'Curved Score', 'Pass/Fail']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12

    # 35 students with realistic names and raw scores (20-95)
    # C and D columns are intentionally EMPTY (task asks to fill them)
    students = [
        ('Aiden Brooks', 72),
        ('Brianna Carter', 85),
        ('Carlos Mendez', 58),
        ('Diana Lee', 91),
        ('Ethan Nguyen', 43),
        ('Fiona Walsh', 67),
        ('Gabriel Torres', 78),
        ('Hannah Kim', 55),
        ('Ivan Petrov', 88),
        ('Jessica Park', 62),
        ('Kevin Martinez', 34),
        ('Laura Chen', 76),
        ('Michael Osei', 49),
        ('Natasha Patel', 83),
        ('Oliver Huang', 70),
        ('Priya Sharma', 92),
        ('Quincy Adams', 27),
        ('Rachel Thompson', 64),
        ('Samuel Wilson', 81),
        ('Tiffany Brown', 56),
        ('Umar Hassan', 38),
        ('Victoria Santos', 74),
        ('William Evans', 89),
        ('Xiaomei Zhang', 47),
        ('Yasmin Ali', 66),
        ('Zoe Robinson', 79),
        ('Aaron Mitchell', 53),
        ('Belinda Flores', 87),
        ('Christopher Diaz', 20),
        ('Danielle Moore', 95),
        ('Edward Taylor', 60),
        ('Florence Jackson', 42),
        ('Gregory White', 73),
        ('Hana Yamamoto', 31),
        ('Isabelle Green', 84),
    ]

    for r, (name, raw_score) in enumerate(students, 2):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=raw_score)
        # Columns C and D are intentionally left empty

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Midterm')
    print(f'  Headers: Student Name, Raw Score, Curved Score, Pass/Fail')
    print(f'  Rows: 35 students (rows 2-36)')
    print(f'  Columns C and D: EMPTY (to be filled by agent)')


create_initial()
