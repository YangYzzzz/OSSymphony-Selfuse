"""
Reward Script: Protect 'Lab Results' sheet with password, unlock B2:B20, restrict selection to unlocked cells
Task ID: calc_ps_018
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): Sheet protection is enabled
  Component 2 (0.35): Cells B2:B20 are unlocked while other cells remain locked
  Component 3 (0.30): Selection restricted to unlocked cells only (selectLockedCells=True)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_018'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Lab Results' sheet must exist
    if 'Lab Results' not in wb.sheetnames:
        print(f"FAIL: 'Lab Results' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Lab Results']

    # Component 1: Sheet protection is enabled (0.35 points)
    # Initial: protection.sheet = False; Golden: protection.sheet = True
    try:
        if ws.protection.sheet:
            print(f"PASS: Component 1 - Sheet protection is enabled (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 1 - Sheet protection is not enabled (protection.sheet={ws.protection.sheet})")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: B2:B20 are unlocked, surrounding cells remain locked (0.35 points)
    # Initial: all cells locked=True; Golden: B2:B20 locked=False, others locked=True
    try:
        unlocked_count = 0
        total_b_cells = 0
        for row in range(2, 21):  # B2 through B20
            cell = ws.cell(row=row, column=2)
            total_b_cells += 1
            if not cell.protection.locked:
                unlocked_count += 1

        # Check that cells outside B2:B20 are still locked
        locked_outside = 0
        total_outside = 0
        check_cells = [
            ('A', 1), ('B', 1), ('C', 2), ('D', 2),  # Header and other columns
            ('A', 2), ('A', 10),  # Column A data cells
            ('B', 21),  # Cell just below B20
        ]
        for col_letter, row_num in check_cells:
            cell = ws[f"{col_letter}{row_num}"]
            total_outside += 1
            if cell.protection.locked:
                locked_outside += 1

        # All 19 B2:B20 cells must be unlocked
        if unlocked_count == 19 and locked_outside == total_outside:
            print(f"PASS: Component 2 - All B2:B20 unlocked ({unlocked_count}/19), outside cells locked ({locked_outside}/{total_outside}) (0.35 pts)")
            total_score += 0.35
        elif unlocked_count == 19:
            # Partial: B2:B20 correct but some outside cells also unlocked
            print(f"PARTIAL: Component 2 - B2:B20 unlocked correctly, but some outside cells also unlocked ({locked_outside}/{total_outside} locked)")
            total_score += 0.20
        elif unlocked_count > 0:
            # Partial: some B cells unlocked
            ratio = unlocked_count / 19
            partial = round(0.35 * ratio * 0.5, 2)  # half credit scaled by ratio
            print(f"PARTIAL: Component 2 - Only {unlocked_count}/19 B2:B20 cells unlocked ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 - No B2:B20 cells are unlocked (all {total_b_cells} still locked)")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Selection restricted to unlocked cells (0.30 points)
    # Initial: selectLockedCells=False; Golden: selectLockedCells=True
    # In openpyxl, selectLockedCells=True means users CANNOT select locked cells
    try:
        select_locked = ws.protection.selectLockedCells
        if select_locked:
            print(f"PASS: Component 3 - selectLockedCells=True (users cannot select locked cells) (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 3 - selectLockedCells={select_locked} (users can still select locked cells)")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
