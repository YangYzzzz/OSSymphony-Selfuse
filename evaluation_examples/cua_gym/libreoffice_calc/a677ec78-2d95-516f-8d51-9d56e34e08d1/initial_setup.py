"""
Initial Setup: Protect Lab Results sheet with unlocked data entry cells
Task ID: calc_ps_018
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_018'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Lab Results'

    # --- Headers ---
    headers = ['Sample ID', 'Measurement', 'Unit', 'Date']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_font_white = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # --- Sample IDs (A2:A20) ---
    sample_ids = [
        'WQ-2025-001', 'WQ-2025-002', 'WQ-2025-003', 'WQ-2025-004',
        'WQ-2025-005', 'WQ-2025-006', 'WQ-2025-007', 'WQ-2025-008',
        'WQ-2025-009', 'WQ-2025-010', 'WQ-2025-011', 'WQ-2025-012',
        'WQ-2025-013', 'WQ-2025-014', 'WQ-2025-015', 'WQ-2025-016',
        'WQ-2025-017', 'WQ-2025-018', 'WQ-2025-019',
    ]
    for r, sid in enumerate(sample_ids, 2):
        cell = ws.cell(row=r, column=1, value=sid)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # --- B2:B20 empty (for data entry) ---
    for r in range(2, 21):
        cell = ws.cell(row=r, column=2)
        cell.border = thin_border
        cell.number_format = '0.00'

    # --- Units (C2:C20) ---
    units = [
        'mg/L', 'mg/L', 'pH', 'mg/L', 'NTU',
        'mg/L', 'ug/L', 'mg/L', 'CFU/100mL', 'mg/L',
        'mg/L', 'pH', 'NTU', 'mg/L', 'ug/L',
        'mg/L', 'CFU/100mL', 'mg/L', 'mg/L',
    ]
    for r, unit in enumerate(units, 2):
        cell = ws.cell(row=r, column=3, value=unit)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # --- Dates (D2:D20) ---
    dates = [
        '2025-03-01', '2025-03-01', '2025-03-02', '2025-03-02', '2025-03-03',
        '2025-03-03', '2025-03-04', '2025-03-04', '2025-03-05', '2025-03-05',
        '2025-03-06', '2025-03-06', '2025-03-07', '2025-03-07', '2025-03-08',
        '2025-03-08', '2025-03-09', '2025-03-09', '2025-03-10',
    ]
    for r, d in enumerate(dates, 2):
        cell = ws.cell(row=r, column=4, value=d)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # --- Column widths ---
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14

    # All cells are locked by default in openpyxl (Protection(locked=True))
    # Sheet is NOT protected - this is the initial state

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
