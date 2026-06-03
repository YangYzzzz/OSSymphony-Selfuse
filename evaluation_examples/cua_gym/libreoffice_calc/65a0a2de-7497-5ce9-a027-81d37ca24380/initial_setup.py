"""
Initial Setup: Marketing Spend spreadsheet with 320 rows of data
Task ID: calc_pivot_055
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_055'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)

    channels = ['Social', 'Search', 'Display', 'Email', 'Content']
    # 320 rows: 64 rows per channel, 16 per channel per quarter
    # We need specific sums: Social/Q1=12000, Search/Q2=18000, Grand total=420000

    # Design the spend distribution so pivot totals are exact.
    # 5 channels x 4 quarters = 20 cells, 16 rows each = 320 rows
    # Grand total = 420000
    # We'll set target sums for each channel/quarter combo

    # Target pivot table (Channel x Quarter -> sum of spend):
    #           Q1      Q2      Q3      Q4    Total
    # Social   12000   14000   16000   13000   55000
    # Search   15000   18000   20000   17000   70000
    # Display  10000   12000   14000   11000   47000
    # Email     8000   10000   12000    9000   39000
    # Content  45000   48000   52000   64000  209000
    # Total    90000  102000  114000  114000  420000

    targets = {
        ('Social', 1): 12000, ('Social', 2): 14000, ('Social', 3): 16000, ('Social', 4): 13000,
        ('Search', 1): 15000, ('Search', 2): 18000, ('Search', 3): 20000, ('Search', 4): 17000,
        ('Display', 1): 10000, ('Display', 2): 12000, ('Display', 3): 14000, ('Display', 4): 11000,
        ('Email', 1): 8000, ('Email', 2): 10000, ('Email', 3): 12000, ('Email', 4): 9000,
        ('Content', 1): 45000, ('Content', 2): 48000, ('Content', 3): 52000, ('Content', 4): 64000,
    }

    # Campaign names per channel
    campaigns = {
        'Social': ['Spring Social Blitz', 'Summer Engagement', 'Fall Community', 'Winter Viral', 'Brand Awareness Social', 'Influencer Partnership', 'Social Retargeting', 'UGC Campaign'],
        'Search': ['Google Ads Primary', 'Bing Ads Growth', 'Product Listing Ads', 'Brand Keywords', 'Competitor Conquest', 'Long Tail SEO Push', 'Local Search', 'Shopping Ads'],
        'Display': ['Banner Retargeting', 'Programmatic Display', 'Native Ads', 'Video Pre-roll', 'Contextual Targeting', 'Publisher Direct', 'Rich Media Ads', 'Interstitial Campaigns'],
        'Email': ['Newsletter Weekly', 'Drip Nurture Series', 'Win-back Campaign', 'Product Launch Email', 'Seasonal Promo', 'Loyalty Program', 'Onboarding Sequence', 'Re-engagement'],
        'Content': ['Blog Syndication', 'Whitepaper Promotion', 'Webinar Series', 'Podcast Sponsorship', 'Case Study Distribution', 'Ebook Launch', 'Industry Report', 'Video Content Series'],
    }

    # Quarter date ranges
    quarter_dates = {
        1: (date(2024, 1, 1), date(2024, 3, 31)),
        2: (date(2024, 4, 1), date(2024, 6, 30)),
        3: (date(2024, 7, 1), date(2024, 9, 30)),
        4: (date(2024, 10, 1), date(2024, 12, 31)),
    }

    def random_date_in_quarter(q):
        start, end = quarter_dates[q]
        delta = (end - start).days
        return start + timedelta(days=random.randint(0, delta))

    rows_per_group = 16  # 5 channels x 4 quarters x 16 = 320

    data_rows = []
    spend_id = 1

    for channel in channels:
        for quarter in range(1, 5):
            target_sum = targets[(channel, quarter)]
            # Generate 16 spend values that sum to target_sum
            # Use random proportions
            raw = [random.random() for _ in range(rows_per_group)]
            raw_sum = sum(raw)
            # Scale to target, round to 2 decimals
            spends = [round(target_sum * r / raw_sum, 2) for r in raw]
            # Adjust last value to ensure exact sum
            spends[-1] = round(target_sum - sum(spends[:-1]), 2)

            for i, spend in enumerate(spends):
                d = random_date_in_quarter(quarter)
                campaign = campaigns[channel][i % len(campaigns[channel])]
                impressions = random.randint(500, 50000)
                clicks = random.randint(int(impressions * 0.005), int(impressions * 0.08))
                data_rows.append((spend_id, d, channel, campaign, spend, impressions, clicks))
                spend_id += 1

    # Shuffle so data isn't grouped by channel/quarter (more realistic)
    random.shuffle(data_rows)
    # Re-assign SpendIDs after shuffle
    data_rows = [(i + 1, r[1], r[2], r[3], r[4], r[5], r[6]) for i, r in enumerate(data_rows)]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'MarketingSpend'

    # Headers
    headers = ['SpendID', 'Date', 'Channel', 'Campaign', 'Spend', 'Impressions', 'Clicks']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # Data rows
    for r, row_data in enumerate(data_rows, 2):
        ws.cell(row=r, column=1, value=row_data[0])  # SpendID
        ws.cell(row=r, column=2, value=row_data[1])   # Date
        ws.cell(row=r, column=2).number_format = 'MM/DD/YYYY'
        ws.cell(row=r, column=3, value=row_data[2])   # Channel
        ws.cell(row=r, column=4, value=row_data[3])   # Campaign
        cell_spend = ws.cell(row=r, column=5, value=row_data[4])  # Spend
        cell_spend.number_format = '$#,##0.00'
        ws.cell(row=r, column=6, value=row_data[5])   # Impressions
        ws.cell(row=r, column=6).number_format = '#,##0'
        ws.cell(row=r, column=7, value=row_data[6])   # Clicks
        ws.cell(row=r, column=7).number_format = '#,##0'

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 10

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
