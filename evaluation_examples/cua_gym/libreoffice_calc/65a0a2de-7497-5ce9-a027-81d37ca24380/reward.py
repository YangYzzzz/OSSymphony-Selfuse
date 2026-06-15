"""
Reward Script: Build a pivot table breaking down marketing spend by channel and quarter
Task ID: calc_pivot_055
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20): PivotTable sheet exists
  Component 2 (0.20): Correct structure (Channel rows, Quarter columns, SUM of Spend header)
  Component 3 (0.15): All 5 channels present as row labels
  Component 4 (0.20): Key spot-check values (Social/Q1=12000, Search/Q2=18000)
  Component 5 (0.15): Grand total = 420000
  Component 6 (0.10): Quarterly column totals correct (90000, 102000, 114000, 114000)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_055'


def persist_app_state(domain: str):
    """Save any unsaved LibreOffice edits before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            import time
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify pivot table creation with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: PivotTable sheet exists (0.20 points)
    # This FAILS on initial (only MarketingSpend), PASSES on golden
    try:
        pivot_sheet_found = False
        pivot_ws = None
        for name in wb.sheetnames:
            if 'pivot' in name.lower():
                pivot_sheet_found = True
                pivot_ws = wb[name]
                break
        if pivot_sheet_found:
            print(f"PASS: Component 1 — Pivot sheet found: '{pivot_ws.title}' (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 — No sheet with 'pivot' in name. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0  # No pivot sheet means nothing else can pass
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 2: Correct structure — Channel rows, Quarter columns, SUM of Spend header (0.20 points)
    # Checks: A1 or A2 contains header indicating data field (sum of spend),
    #         Row 2 has quarter labels, Column A has channel names
    try:
        structure_score = 0.0

        # Check for "Spend" reference in header area (A1)
        a1_val = str(pivot_ws.cell(row=1, column=1).value or '').lower()
        has_spend_header = 'spend' in a1_val
        if has_spend_header:
            structure_score += 0.05
            print(f"  PASS: A1 contains spend reference: '{pivot_ws.cell(row=1, column=1).value}'")
        else:
            print(f"  FAIL: A1 does not reference 'spend': '{pivot_ws.cell(row=1, column=1).value}'")

        # Check row 2 has "Channel" label in A2
        a2_val = str(pivot_ws.cell(row=2, column=1).value or '').lower()
        has_channel_label = 'channel' in a2_val
        if has_channel_label:
            structure_score += 0.05
            print(f"  PASS: A2 contains 'Channel': '{pivot_ws.cell(row=2, column=1).value}'")
        else:
            print(f"  FAIL: A2 does not contain 'Channel': '{pivot_ws.cell(row=2, column=1).value}'")

        # Check quarter columns exist in row 2 (B2 onwards)
        quarter_labels = []
        for col in range(2, pivot_ws.max_column + 1):
            val = str(pivot_ws.cell(row=2, column=col).value or '').lower()
            quarter_labels.append(val)
        q_count = sum(1 for q in quarter_labels if 'q' in q and '2024' in q)
        if q_count >= 4:
            structure_score += 0.05
            print(f"  PASS: Found {q_count} quarter columns in row 2")
        else:
            print(f"  FAIL: Expected 4 quarter columns, found {q_count}. Labels: {quarter_labels}")

        # Check Grand Total column exists
        has_grand_total_col = any('grand' in q and 'total' in q for q in quarter_labels)
        if has_grand_total_col:
            structure_score += 0.05
            print(f"  PASS: Grand Total column found")
        else:
            print(f"  FAIL: No Grand Total column found. Labels: {quarter_labels}")

        if structure_score > 0:
            print(f"PASS: Component 2 — Structure score: {structure_score:.2f}/0.20")
            total_score += structure_score
        else:
            print(f"FAIL: Component 2 — No structural elements matched")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: All 5 channels present as row labels (0.15 points)
    # Expected channels: Content, Display, Email, Search, Social
    try:
        expected_channels = {'content', 'display', 'email', 'search', 'social'}
        found_channels = set()
        for row in range(3, pivot_ws.max_row + 1):
            val = pivot_ws.cell(row=row, column=1).value
            if val and str(val).lower().strip() in expected_channels:
                found_channels.add(str(val).lower().strip())

        if found_channels == expected_channels:
            print(f"PASS: Component 3 — All 5 channels found: {sorted(found_channels)} (0.15 pts)")
            total_score += 0.15
        else:
            missing = expected_channels - found_channels
            print(f"FAIL: Component 3 — Missing channels: {missing}. Found: {sorted(found_channels)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Helper: find the column index for a given quarter label in row 2
    def find_col(label_substr):
        for col in range(2, pivot_ws.max_column + 1):
            val = str(pivot_ws.cell(row=2, column=col).value or '').lower()
            if label_substr.lower() in val:
                return col
        return None

    # Helper: find the row index for a given channel name
    def find_row(channel_name):
        for row in range(3, pivot_ws.max_row + 1):
            val = str(pivot_ws.cell(row=row, column=1).value or '').lower().strip()
            if val == channel_name.lower():
                return row
        return None

    # Component 4: Key spot-check values (0.20 points)
    # Social/Q1 = 12000, Search/Q2 = 18000
    try:
        spot_score = 0.0

        # Social / Q1 = 12000
        social_row = find_row('Social')
        q1_col = find_col('q1')
        if social_row and q1_col:
            social_q1_val = pivot_ws.cell(row=social_row, column=q1_col).value
            try:
                if social_q1_val is not None and abs(float(social_q1_val) - 12000) < 1:
                    spot_score += 0.10
                    print(f"  PASS: Social/Q1 = {social_q1_val} (expected 12000)")
                else:
                    print(f"  FAIL: Social/Q1 = {social_q1_val} (expected 12000)")
            except (ValueError, TypeError):
                print(f"  FAIL: Social/Q1 not numeric: {social_q1_val}")
        else:
            print(f"  FAIL: Could not locate Social row ({social_row}) or Q1 col ({q1_col})")

        # Search / Q2 = 18000
        search_row = find_row('Search')
        q2_col = find_col('q2')
        if search_row and q2_col:
            search_q2_val = pivot_ws.cell(row=search_row, column=q2_col).value
            try:
                if search_q2_val is not None and abs(float(search_q2_val) - 18000) < 1:
                    spot_score += 0.10
                    print(f"  PASS: Search/Q2 = {search_q2_val} (expected 18000)")
                else:
                    print(f"  FAIL: Search/Q2 = {search_q2_val} (expected 18000)")
            except (ValueError, TypeError):
                print(f"  FAIL: Search/Q2 not numeric: {search_q2_val}")
        else:
            print(f"  FAIL: Could not locate Search row ({search_row}) or Q2 col ({q2_col})")

        if spot_score > 0:
            print(f"PASS: Component 4 — Spot-check score: {spot_score:.2f}/0.20")
            total_score += spot_score
        else:
            print(f"FAIL: Component 4 — No spot-check values matched")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Grand total = 420000 (0.15 points)
    try:
        grand_total_found = False
        # Look for "Grand Total" row and its last numeric column
        for row in range(3, pivot_ws.max_row + 1):
            val = str(pivot_ws.cell(row=row, column=1).value or '').lower().strip()
            if 'grand' in val and 'total' in val:
                # Find the Grand Total column (last column with data)
                gt_col = find_col('grand total')
                if gt_col is None:
                    gt_col = pivot_ws.max_column  # fallback to last column
                gt_val = pivot_ws.cell(row=row, column=gt_col).value
                try:
                    if gt_val is not None and abs(float(gt_val) - 420000) < 1:
                        grand_total_found = True
                        print(f"PASS: Component 5 — Grand Total = {gt_val} (expected 420000) (0.15 pts)")
                        total_score += 0.15
                    else:
                        print(f"FAIL: Component 5 — Grand Total = {gt_val} (expected 420000)")
                except (ValueError, TypeError):
                    print(f"FAIL: Component 5 — Grand Total not numeric: {gt_val}")
                break
        if not grand_total_found and total_score > 0.2:
            print(f"FAIL: Component 5 — No 'Grand Total' row found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Quarterly column totals correct (0.10 points)
    # Expected: Q1=90000, Q2=102000, Q3=114000, Q4=114000
    try:
        expected_totals = {'q1': 90000, 'q2': 102000, 'q3': 114000, 'q4': 114000}
        correct_count = 0

        # Find the Grand Total row
        gt_row = None
        for row in range(3, pivot_ws.max_row + 1):
            val = str(pivot_ws.cell(row=row, column=1).value or '').lower().strip()
            if 'grand' in val and 'total' in val:
                gt_row = row
                break

        if gt_row:
            for q_label, expected_val in expected_totals.items():
                q_col = find_col(q_label)
                if q_col:
                    actual_val = pivot_ws.cell(row=gt_row, column=q_col).value
                    try:
                        if actual_val is not None and abs(float(actual_val) - expected_val) < 1:
                            correct_count += 1
                            print(f"  PASS: {q_label.upper()} total = {actual_val} (expected {expected_val})")
                        else:
                            print(f"  FAIL: {q_label.upper()} total = {actual_val} (expected {expected_val})")
                    except (ValueError, TypeError):
                        print(f"  FAIL: {q_label.upper()} total not numeric: {actual_val}")
                else:
                    print(f"  FAIL: Column for {q_label.upper()} not found")

            if correct_count == 4:
                print(f"PASS: Component 6 — All 4 quarterly totals correct (0.10 pts)")
                total_score += 0.10
            elif correct_count > 0:
                partial = round(0.10 * correct_count / 4, 2)
                print(f"PARTIAL: Component 6 — {correct_count}/4 quarterly totals correct ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 6 — No quarterly totals matched")
        else:
            print(f"FAIL: Component 6 — No Grand Total row found to check quarterly totals")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
