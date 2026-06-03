"""
FINAL REWARD SCRIPT - SUCCESS
Task: Please copy the 'Product SKU' column to 'Full SKU Code' and pad each entry with leading zeros to make them 6 digits. Don't modify any blank or unrelated cells.
Generated: 2025-11-24 07:25:36
Status: success
Model: o3
Total Steps: 1
"""

import openpyxl
import os

# -----------------------------------------------------------------------------
# Reward Verification Script
# -----------------------------------------------------------------------------
# Task:  Copy the values from column 'Product SKU' to column 'Full SKU Code'
#        and pad each copied value on the left with zeros so it becomes a
#        6-character string.  All other data (headers, descriptions, etc.) must
#        remain unchanged.
# -----------------------------------------------------------------------------
# Scoring philosophy
#   •  0.75  – correct "Full SKU Code" for each of the 5 data rows (0.15 each)
#   •  0.10  – original "Product SKU" values remain unchanged
#   •  0.05  – "Description" values remain unchanged
#   •  0.05  – header row remains unchanged
#   •  0.05  – row count / basic structure intact
#   = 1.00   – PERFECT SCORE
# -----------------------------------------------------------------------------
# IMPORTANT:  No points are awarded for natural conditions (file exists, loads,
#             etc.).  All points are earned only for meeting specific task
#             requirements.
# -----------------------------------------------------------------------------

def verify_padding_task(file_path: str) -> float:
    """Verify that the spreadsheet satisfies the task requirements.

    Returns
    -------
    float
        A score between 0.0 and 1.0 indicating task completion.
    """

    print(f"Verifying task for file: {file_path}")
    total_score = 0.0
    MAX_SCORE  = 1.0

    # ---------------------------------------------------------------------
    # Load workbook
    # ---------------------------------------------------------------------
    try:
        wb = openpyxl.load_workbook(file_path)
        print("✓ Workbook loaded successfully")
    except Exception as e:
        print(f"✗ Unable to load workbook: {e}")
        return 0.0

    # ---------------------------------------------------------------------
    # Locate required sheet & columns
    # ---------------------------------------------------------------------
    SHEET_NAME = "Sheet"
    if SHEET_NAME not in wb.sheetnames:
        print(f"✗ Required sheet '{SHEET_NAME}' not found")
        return 0.0
    ws = wb[SHEET_NAME]

    header_row = [cell.value for cell in ws[1]]
    try:
        col_product = header_row.index("Product SKU") + 1  # 1-based index
        col_full    = header_row.index("Full SKU Code") + 1
        col_desc    = header_row.index("Description") + 1
        print(f"✓ Column indices – Product SKU: {col_product}, Full SKU Code: {col_full}, Description: {col_desc}")
    except ValueError:
        print("✗ One or more required headers are missing or renamed")
        return 0.0

    # ---------------------------------------------------------------------
    # Expected original data (from task description)
    # ---------------------------------------------------------------------
    original_skus  = ["100", "5235", "98765", "42", "700001"]
    original_descs = ["Widget A", "Widget B", "Widget C", "Widget D", "Widget E"]

    # ---------------------------------------------------------------------
    # Helper – normalise cell value to trimmed string
    # ---------------------------------------------------------------------
    def norm(value):
        if value is None:
            return ""
        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            return str(value)
        return str(value).strip()

    # ---------------------------------------------------------------------
    # 1) Header unchanged (0.05)
    # ---------------------------------------------------------------------
    expected_headers = ["Product SKU", "Full SKU Code", "Description"]
    if all(h1 == h2 for h1, h2 in zip(expected_headers, header_row)):
        total_score += 0.05
        print("✓ Header row unchanged (0.05)")
    else:
        print("✗ Header row was modified (0 points)")

    # ---------------------------------------------------------------------
    # 2) Row count / basic structure intact (0.05)
    # ---------------------------------------------------------------------
    if ws.max_row >= 6:
        total_score += 0.05
        print(f"✓ Expected row count present (≥6 rows) (0.05)")
    else:
        print(f"✗ Row count too small (found {ws.max_row}) (0 points)")

    # ---------------------------------------------------------------------
    # 3) Per-row checks (rows 2-6)
    # ---------------------------------------------------------------------
    full_sku_correct = 0
    product_ok       = True
    desc_ok          = True

    for idx, excel_row in enumerate(range(2, 7)):
        raw_product   = norm(ws.cell(row=excel_row, column=col_product).value)
        raw_fullsku   = norm(ws.cell(row=excel_row, column=col_full).value)
        raw_desc      = ws.cell(row=excel_row, column=col_desc).value

        expected_product = original_skus[idx]
        expected_desc    = original_descs[idx]
        expected_full    = expected_product.zfill(6) if len(expected_product) <= 6 else expected_product

        # Product SKU column must remain unchanged
        if raw_product != expected_product:
            product_ok = False
            print(f"✗ Row {excel_row}: Product SKU altered (exp='{expected_product}', found='{raw_product}')")

        # Description column must remain unchanged
        if raw_desc != expected_desc:
            desc_ok = False
            print(f"✗ Row {excel_row}: Description altered (exp='{expected_desc}', found='{raw_desc}')")

        # Full SKU Code must be correctly copied & padded
        if raw_fullsku == expected_full:
            full_sku_correct += 1
            print(f"✓ Row {excel_row}: Full SKU Code correct ('{raw_fullsku}')")
        else:
            print(f"✗ Row {excel_row}: Full SKU Code incorrect (exp='{expected_full}', found='{raw_fullsku}')")

    # ---------------------------------------------------------------------
    # 3a) Scoring for Full SKU Code correctness (0.15 each)
    # ---------------------------------------------------------------------
    total_score += full_sku_correct * 0.15

    # ---------------------------------------------------------------------
    # 3b) Additional column integrity scores
    # ---------------------------------------------------------------------
    if product_ok:
        total_score += 0.10
        print("✓ Product SKU column unchanged (0.10)")
    else:
        print("✗ Product SKU column has changes (0 points)")

    if desc_ok:
        total_score += 0.05
        print("✓ Description column unchanged (0.05)")
    else:
        print("✗ Description column has changes (0 points)")

    # ---------------------------------------------------------------------
    # Final score (capped and rounded)
    # ---------------------------------------------------------------------
    final_score = round(min(total_score, MAX_SCORE), 4)
    print(f"Total score: {final_score}/{MAX_SCORE}")
    return final_score

# -----------------------------------------------------------------------------
# MAIN EXECUTION (Called when script is run)
# -----------------------------------------------------------------------------
if __name__ == "__main__":
    FILE_PATH = "/home/user/please_copy_the_product_sku_column_to_full_sku_code_and_pad_each_entry_with_leading_zeros_to_make_th.xlsx"

    if not os.path.isfile(FILE_PATH):
        print(f"✗ File not found: {FILE_PATH}")
        print("REWARD: 0.0")
    else:
        reward = verify_padding_task(FILE_PATH)
        print(f"REWARD: {reward}")
