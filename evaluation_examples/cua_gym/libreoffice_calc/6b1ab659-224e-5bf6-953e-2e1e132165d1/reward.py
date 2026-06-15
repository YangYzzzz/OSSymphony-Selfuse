"""
Reward Script: Apply borders to range A1:F15 in schedule.xlsx
Task ID: calc_gg5_010
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): Corner cells have correct thick outer + thin inner borders
  Component 2 (0.35): Edge cells have correct medium outer edge + thin inner borders
  Component 3 (0.30): Interior cells have all-thin borders
  Precondition gate: Cells outside A1:F15 have no borders (no points, early exit if violated)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_010'


def persist_app_state(domain: str):
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Shifts' sheet must exist
    if 'Shifts' not in wb.sheetnames:
        print("CRITICAL: 'Shifts' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Shifts']

    # Helper: check if a border side has a "thick" outer style (medium or thick)
    def is_outer_style(style):
        return style in ('medium', 'thick')

    # Helper: check if a border side is thin
    def is_thin(style):
        return style == 'thin'

    # Helper: check if a border side is None/absent
    def is_none(style):
        return style is None

    # Component 1: Corner cells have correct outer borders (0.3 points)
    # A1: top=medium, left=medium, right=thin, bottom=thin
    # F1: top=medium, right=medium, left=thin, bottom=thin
    # A15: bottom=medium, left=medium, right=thin, top=thin
    # F15: bottom=medium, right=medium, left=thin, top=thin
    try:
        corners_pass = 0
        corner_checks = [
            ('A1',  {'top': 'outer', 'left': 'outer', 'right': 'thin', 'bottom': 'thin'}),
            ('F1',  {'top': 'outer', 'right': 'outer', 'left': 'thin', 'bottom': 'thin'}),
            ('A15', {'bottom': 'outer', 'left': 'outer', 'right': 'thin', 'top': 'thin'}),
            ('F15', {'bottom': 'outer', 'right': 'outer', 'left': 'thin', 'top': 'thin'}),
        ]
        for coord, expected in corner_checks:
            b = ws[coord].border
            sides = {'top': b.top.style, 'bottom': b.bottom.style,
                     'left': b.left.style, 'right': b.right.style}
            ok = True
            for side_name, req in expected.items():
                actual = sides[side_name]
                if req == 'outer':
                    if not is_outer_style(actual):
                        ok = False
                elif req == 'thin':
                    if not is_thin(actual):
                        ok = False
            if ok:
                corners_pass += 1
                print(f"PASS: Corner {coord} has correct borders")
            else:
                print(f"FAIL: Corner {coord} — expected {expected}, got L={sides['left']} R={sides['right']} T={sides['top']} B={sides['bottom']}")

        if corners_pass == 4:
            print(f"PASS: Component 1 — All 4 corners correct (0.35 pts)")
            total_score += 0.35
        elif corners_pass >= 2:
            partial = round(0.35 * corners_pass / 4, 2)
            print(f"PARTIAL: Component 1 — {corners_pass}/4 corners correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {corners_pass}/4 corners correct (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Edge cells (non-corner, on border of A1:F15) have correct borders (0.3 points)
    # Top edge (B1-E1): top=medium, bottom=thin, left=thin, right=thin
    # Bottom edge (B15-E15): bottom=medium, top=thin, left=thin, right=thin
    # Left edge (A2-A14): left=medium, right=thin, top=thin, bottom=thin
    # Right edge (F2-F14): right=medium, left=thin, top=thin, bottom=thin
    try:
        edge_total = 0
        edge_pass = 0

        # Top edge: B1 to E1
        for col in range(2, 6):  # B=2, E=5
            coord = ws.cell(row=1, column=col)
            b = coord.border
            edge_total += 1
            if is_outer_style(b.top.style) and is_thin(b.bottom.style) and is_thin(b.left.style) and is_thin(b.right.style):
                edge_pass += 1

        # Bottom edge: B15 to E15
        for col in range(2, 6):
            coord = ws.cell(row=15, column=col)
            b = coord.border
            edge_total += 1
            if is_outer_style(b.bottom.style) and is_thin(b.top.style) and is_thin(b.left.style) and is_thin(b.right.style):
                edge_pass += 1

        # Left edge: A2 to A14
        for row in range(2, 15):
            coord = ws.cell(row=row, column=1)
            b = coord.border
            edge_total += 1
            if is_outer_style(b.left.style) and is_thin(b.right.style) and is_thin(b.top.style) and is_thin(b.bottom.style):
                edge_pass += 1

        # Right edge: F2 to F14
        for row in range(2, 15):
            coord = ws.cell(row=row, column=6)
            b = coord.border
            edge_total += 1
            if is_outer_style(b.right.style) and is_thin(b.left.style) and is_thin(b.top.style) and is_thin(b.bottom.style):
                edge_pass += 1

        edge_ratio = edge_pass / edge_total if edge_total > 0 else 0
        if edge_ratio >= 0.9:
            print(f"PASS: Component 2 — {edge_pass}/{edge_total} edge cells correct (0.35 pts)")
            total_score += 0.35
        elif edge_ratio >= 0.5:
            partial = round(0.35 * edge_ratio, 2)
            print(f"PARTIAL: Component 2 — {edge_pass}/{edge_total} edge cells correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — {edge_pass}/{edge_total} edge cells correct (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Interior cells (B2:E14) have all-thin borders (0.2 points)
    try:
        interior_total = 0
        interior_pass = 0
        for row in range(2, 15):
            for col in range(2, 6):  # B=2, E=5
                cell = ws.cell(row=row, column=col)
                b = cell.border
                interior_total += 1
                if is_thin(b.left.style) and is_thin(b.right.style) and is_thin(b.top.style) and is_thin(b.bottom.style):
                    interior_pass += 1

        interior_ratio = interior_pass / interior_total if interior_total > 0 else 0
        if interior_ratio >= 0.9:
            print(f"PASS: Component 3 — {interior_pass}/{interior_total} interior cells have thin borders (0.3 pts)")
            total_score += 0.3
        elif interior_ratio >= 0.5:
            partial = round(0.3 * interior_ratio, 2)
            print(f"PARTIAL: Component 3 — {interior_pass}/{interior_total} interior cells correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — {interior_pass}/{interior_total} interior cells have thin borders (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Precondition gate: Cells outside A1:F15 have NO borders
    # This is true in both initial and golden, so it is NOT a scoring component.
    # Used only as a sanity check — if borders leaked outside, log a warning.
    try:
        outside_coords = ['G1', 'G8', 'G15', 'A16', 'F16', 'A17', 'H1', 'H5']
        outside_clean = 0
        for coord in outside_coords:
            cell = ws[coord]
            b = cell.border
            if is_none(b.left.style) and is_none(b.right.style) and is_none(b.top.style) and is_none(b.bottom.style):
                outside_clean += 1
        if outside_clean == len(outside_coords):
            print(f"GATE: No borders leaked outside A1:F15 ({outside_clean}/{len(outside_coords)} clean)")
        else:
            print(f"GATE_WARN: {len(outside_coords) - outside_clean}/{len(outside_coords)} outside cells have unexpected borders")
    except Exception as e:
        print(f"GATE_WARN: Could not check outside borders — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
