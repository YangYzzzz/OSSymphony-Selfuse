"""
Initial Setup: Apply borders to a staff schedule spreadsheet
Task ID: calc_gg5_010
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_010'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Shifts'

    # Headers (row 1)
    headers = ['Employee', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Staff schedule data (rows 2-15, 14 employees)
    schedule_data = [
        ['Sarah Chen',       '9AM-5PM',  '9AM-5PM',  'OFF',      '9AM-5PM',  '9AM-5PM'],
        ['Marcus Johnson',   '7AM-3PM',  '7AM-3PM',  '7AM-3PM',  '7AM-3PM',  'OFF'],
        ['Priya Patel',      '2PM-10PM', 'OFF',       '2PM-10PM', '2PM-10PM', '2PM-10PM'],
        ['James O\'Brien',   'OFF',       '9AM-5PM',  '9AM-5PM',  '9AM-5PM',  '9AM-5PM'],
        ['Aisha Williams',   '7AM-3PM',  '7AM-3PM',  '7AM-3PM',  'OFF',       '7AM-3PM'],
        ['David Kim',        '9AM-5PM',  '2PM-10PM', '2PM-10PM', '9AM-5PM',  'OFF'],
        ['Elena Rodriguez',  '2PM-10PM', '2PM-10PM', 'OFF',       '2PM-10PM', '9AM-5PM'],
        ['Thomas Wright',    'OFF',       '7AM-3PM',  '9AM-5PM',  '7AM-3PM',  '7AM-3PM'],
        ['Fatima Al-Rashid', '9AM-5PM',  'OFF',       '7AM-3PM',  '9AM-5PM',  '9AM-5PM'],
        ['Robert Chang',     '7AM-3PM',  '9AM-5PM',  '9AM-5PM',  'OFF',       '2PM-10PM'],
        ['Lisa Nakamura',    '2PM-10PM', '2PM-10PM', '2PM-10PM', '7AM-3PM',  'OFF'],
        ['Michael Torres',   '9AM-5PM',  '7AM-3PM',  'OFF',       '2PM-10PM', '9AM-5PM'],
        ['Hannah Osei',      '7AM-3PM',  'OFF',       '7AM-3PM',  '9AM-5PM',  '7AM-3PM'],
        ['Ryan Gupta',       'OFF',       '9AM-5PM',  '2PM-10PM', '7AM-3PM',  '2PM-10PM'],
    ]

    for r, row_data in enumerate(schedule_data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 20
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col_letter].width = 12

    # NO borders applied - this is the initial state

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
