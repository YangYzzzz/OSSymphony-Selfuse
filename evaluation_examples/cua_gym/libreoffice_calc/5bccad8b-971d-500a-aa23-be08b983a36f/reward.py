"""
Reward Script: Pivot table from healthcare data — patient count by diagnosis & age group with gender filter
Task ID: calc_pivot_084
Domain: libreoffice_calc
Scoring:
  Component 1 (0.15): PivotTable sheet exists
  Component 2 (0.25): Correct row/column structure (Diagnosis rows, AgeGroup columns)
  Component 3 (0.15): Gender page/filter field present
  Component 4 (0.15): Data field is Count of PatientID
  Component 5 (0.30): Data values correct (grand total=400, spot checks)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_084'


def persist_app_state(domain):
    """Save any unsaved LibreOffice edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for", domain)
    except Exception as e:
        print("PERSIST_WARN: save hook failed:", e)


def find_pivot_sheet(wb):
    """Find a sheet that looks like a pivot table (not PatientData).
    Returns the worksheet or None.
    """
    for name in wb.sheetnames:
        if name.lower() != 'patientdata':
            return wb[name]
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: PivotTable sheet exists (0.15 points)
    # The golden file has a second sheet for the pivot table; initial has only PatientData
    try:
        pivot_ws = find_pivot_sheet(wb)
        if pivot_ws is not None:
            print(f"PASS: Component 1 — Pivot sheet found: '{pivot_ws.title}' (0.15 pts)")
            total_score += 0.15
        else:
            print("FAIL: Component 1 — No pivot table sheet found (only PatientData)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # If no pivot sheet, remaining components cannot pass
    if pivot_ws is None:
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    # Component 2: Correct structure — Diagnosis as row labels, AgeGroup as column headers (0.25 points)
    # Expected row labels in column A: the 5 diagnosis categories
    # Expected column headers: the 5 age group categories
    try:
        expected_diagnoses = {'cardiac', 'general', 'neurological', 'orthopedic', 'respiratory'}
        expected_age_groups = {'0-17', '18-34', '35-49', '50-64', '65+'}

        # Scan for diagnosis row labels (look in column A, rows 1-20)
        found_diagnoses = set()
        for r in range(1, min(pivot_ws.max_row + 1, 25)):
            val = pivot_ws.cell(row=r, column=1).value
            if val and str(val).strip().lower() in expected_diagnoses:
                found_diagnoses.add(str(val).strip().lower())

        # Scan for age group column headers (look in rows 1-5, columns A-J)
        found_age_groups = set()
        for r in range(1, 6):
            for c in range(1, 11):
                val = pivot_ws.cell(row=r, column=c).value
                if val and str(val).strip() in expected_age_groups:
                    found_age_groups.add(str(val).strip())

        diag_ok = found_diagnoses == expected_diagnoses
        age_ok = found_age_groups == expected_age_groups

        if diag_ok and age_ok:
            print(f"PASS: Component 2 — All 5 diagnosis rows and 5 age group columns found (0.25 pts)")
            total_score += 0.25
        else:
            missing_diag = expected_diagnoses - found_diagnoses
            missing_age = expected_age_groups - found_age_groups
            print(f"FAIL: Component 2 — Missing diagnoses: {missing_diag}, Missing age groups: {missing_age}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Gender page/filter field present (0.15 points)
    # The golden file has "Gender (All)" in A1 as a page filter indicator
    try:
        gender_cells = [
            pivot_ws.cell(row=r, column=c).value
            for r in range(1, 4)
            for c in range(1, 5)
        ]
        gender_filter_found = any(v and 'gender' in str(v).lower() for v in gender_cells)

        if gender_filter_found:
            print(f"PASS: Component 3 — Gender filter/page field found (0.15 pts)")
            total_score += 0.15
        else:
            print("FAIL: Component 3 — No Gender filter/page field reference found in pivot header area")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Data field is Count of PatientID (0.15 points)
    # Look for "Count of PatientID" or similar label in the pivot table
    try:
        label_cells = [
            pivot_ws.cell(row=r, column=c).value
            for r in range(1, 6)
            for c in range(1, 5)
        ]
        count_label_found = any(
            v and 'count' in str(v).lower() and 'patient' in str(v).lower()
            for v in label_cells
        )

        if count_label_found:
            print(f"PASS: Component 4 — 'Count of PatientID' label found (0.15 pts)")
            total_score += 0.15
        else:
            print("FAIL: Component 4 — No 'Count of PatientID' label found in pivot header")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Data values correct (0.30 points)
    # Check grand total = 400, plus spot-check specific cells
    # From golden: Grand Total row G column = 400
    # Row totals: Cardiac=88, General=92, Neurological=62, Orthopedic=82, Respiratory=76
    try:
        sub_score = 0.0

        # Find the grand total cell: look for "Grand Total" in column A
        grand_total_row = None
        for r in range(1, pivot_ws.max_row + 1):
            val = pivot_ws.cell(row=r, column=1).value
            if val and 'grand total' in str(val).lower():
                grand_total_row = r
                break

        # Find the "Grand Total" column header
        grand_total_col = None
        header_row = None
        for r in range(1, 6):
            for c in range(1, pivot_ws.max_column + 1):
                val = pivot_ws.cell(row=r, column=c).value
                if val and 'grand total' in str(val).lower():
                    grand_total_col = c
                    header_row = r
                    break
            if grand_total_col:
                break

        # Check overall grand total = 400 (0.15 pts)
        if grand_total_row and grand_total_col:
            gt_val = pivot_ws.cell(row=grand_total_row, column=grand_total_col).value
            if gt_val is not None:
                try:
                    if abs(float(gt_val) - 400) < 1:
                        print(f"PASS: Component 5a — Grand total = {gt_val} (expected 400) (0.15 pts)")
                        sub_score += 0.15
                    else:
                        print(f"FAIL: Component 5a — Grand total = {gt_val}, expected 400")
                except (ValueError, TypeError):
                    print(f"FAIL: Component 5a — Grand total not numeric: {gt_val}")
            else:
                print("FAIL: Component 5a — Grand total cell is empty")
        else:
            print(f"FAIL: Component 5a — Could not locate Grand Total (row={grand_total_row}, col={grand_total_col})")

        # Spot-check row totals (0.15 pts)
        # Expected: Cardiac=88, General=92, Neurological=62, Orthopedic=82, Respiratory=76
        expected_row_totals = {
            'cardiac': 88,
            'general': 92,
            'neurological': 62,
            'orthopedic': 82,
            'respiratory': 76,
        }

        if grand_total_col and header_row:
            correct_count = 0
            for r in range(header_row + 1, pivot_ws.max_row + 1):
                row_label = pivot_ws.cell(row=r, column=1).value
                if row_label and str(row_label).strip().lower() in expected_row_totals:
                    expected = expected_row_totals[str(row_label).strip().lower()]
                    actual = pivot_ws.cell(row=r, column=grand_total_col).value
                    if actual is not None:
                        try:
                            if abs(float(actual) - expected) < 1:
                                correct_count += 1
                            else:
                                print(f"  INFO: Row total for {row_label}: {actual} (expected {expected})")
                        except (ValueError, TypeError):
                            print(f"  INFO: Row total for {row_label} not numeric: {actual}")

            if correct_count >= 4:
                print(f"PASS: Component 5b — {correct_count}/5 row totals correct (0.15 pts)")
                sub_score += 0.15
            else:
                print(f"FAIL: Component 5b — Only {correct_count}/5 row totals correct")
        else:
            print("FAIL: Component 5b — Cannot check row totals without Grand Total column")

        if sub_score > 0:
            total_score += sub_score
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
