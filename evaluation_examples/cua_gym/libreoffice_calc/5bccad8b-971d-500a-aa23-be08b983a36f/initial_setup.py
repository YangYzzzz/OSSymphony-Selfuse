"""
Initial Setup: Healthcare patient data for pivot table generation
Task ID: calc_pivot_084
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_084'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'PatientData'

    # --- Headers ---
    headers = ['PatientID', 'Age', 'AgeGroup', 'Gender', 'Diagnosis', 'LOS']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Data generation ---
    age_groups = ['0-17', '18-34', '35-49', '50-64', '65+']
    age_ranges = {
        '0-17': (1, 17),
        '18-34': (18, 34),
        '35-49': (35, 49),
        '50-64': (50, 64),
        '65+': (65, 92),
    }
    genders = ['M', 'F']
    diagnoses = ['Cardiac', 'Respiratory', 'Orthopedic', 'Neurological', 'General']

    # Distribution weights to ensure:
    # - ~205 Male, ~195 Female
    # - Cardiac/65+ is the highest cell
    # We'll use weighted sampling for age groups and diagnoses
    # to create a realistic healthcare dataset

    # Age group distribution (more elderly patients for healthcare realism)
    age_group_weights = {
        '0-17': 30,
        '18-34': 55,
        '35-49': 75,
        '50-64': 100,
        '65+': 140,
    }

    # Diagnosis weights vary by age group to make Cardiac/65+ highest
    diagnosis_weights_by_age = {
        '0-17':  {'Cardiac': 2, 'Respiratory': 8, 'Orthopedic': 6, 'Neurological': 4, 'General': 10},
        '18-34': {'Cardiac': 4, 'Respiratory': 10, 'Orthopedic': 14, 'Neurological': 8, 'General': 19},
        '35-49': {'Cardiac': 12, 'Respiratory': 14, 'Orthopedic': 18, 'Neurological': 12, 'General': 19},
        '50-64': {'Cardiac': 28, 'Respiratory': 18, 'Orthopedic': 20, 'Neurological': 16, 'General': 18},
        '65+':   {'Cardiac': 42, 'Respiratory': 26, 'Orthopedic': 24, 'Neurological': 22, 'General': 26},
    }

    # LOS ranges by diagnosis (realistic hospital stays)
    los_ranges = {
        'Cardiac': (2, 14),
        'Respiratory': (1, 10),
        'Orthopedic': (1, 7),
        'Neurological': (3, 21),
        'General': (1, 5),
    }

    # Build the full population list based on weights
    population = []
    for ag, ag_count in age_group_weights.items():
        diag_w = diagnosis_weights_by_age[ag]
        for diag, d_count in diag_w.items():
            for _ in range(d_count):
                population.append((ag, diag))

    # Shuffle and take 400
    random.shuffle(population)
    population = population[:400]

    # Assign genders: ~205 M, ~195 F
    gender_list = ['M'] * 205 + ['F'] * 195
    random.shuffle(gender_list)

    rows = []
    for i, ((ag, diag), gender) in enumerate(zip(population, gender_list)):
        pid = f'P{i+1:03d}'
        lo, hi = age_ranges[ag]
        age = random.randint(lo, hi)
        los_lo, los_hi = los_ranges[diag]
        los = random.randint(los_lo, los_hi)
        rows.append([pid, age, ag, gender, diag, los])

    # Sort by PatientID for clean presentation
    rows.sort(key=lambda r: r[0])

    # Write data
    for r_idx, row_data in enumerate(rows, 2):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            if c_idx in (1, 3, 4, 5):  # Center text columns
                cell.alignment = Alignment(horizontal='center')

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 8

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Total rows: {len(rows)}')

    # Count for verification
    from collections import Counter
    gender_count = Counter(r[3] for r in rows)
    print(f'Gender distribution: {dict(gender_count)}')
    diag_age = Counter((r[4], r[2]) for r in rows)
    print(f'Cardiac/65+: {diag_age.get(("Cardiac", "65+"), 0)}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
