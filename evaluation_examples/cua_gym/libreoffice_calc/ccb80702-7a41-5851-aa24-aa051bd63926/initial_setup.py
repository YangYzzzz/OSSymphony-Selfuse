"""
Initial Setup: Student Performance Tracker Dashboard
Task ID: calc_gsd_041
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_041'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tracker"

    # Headers in row 1
    headers = ["Student ID", "Name", "Attendance%", "HW Average",
               "Test Average", "Final Grade", "At Risk"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 40 students with realistic data
    first_names = [
        "Sarah", "Marcus", "Olivia", "James", "Emma", "Liam", "Sophia", "Noah",
        "Isabella", "Ethan", "Mia", "Lucas", "Charlotte", "Mason", "Amelia",
        "Logan", "Harper", "Alexander", "Evelyn", "Benjamin", "Abigail", "Daniel",
        "Emily", "Henry", "Elizabeth", "Sebastian", "Avery", "Jack", "Ella",
        "Owen", "Scarlett", "Samuel", "Grace", "Ryan", "Chloe", "Nathan",
        "Victoria", "Caleb", "Riley", "Dylan"
    ]
    last_names = [
        "Chen", "Johnson", "Williams", "Garcia", "Martinez", "Anderson", "Taylor",
        "Thomas", "Hernandez", "Moore", "Martin", "Jackson", "Thompson", "White",
        "Lopez", "Lee", "Gonzalez", "Harris", "Clark", "Lewis", "Robinson",
        "Walker", "Perez", "Hall", "Young", "Allen", "Sanchez", "Wright",
        "King", "Scott", "Green", "Baker", "Adams", "Nelson", "Hill",
        "Ramirez", "Campbell", "Mitchell", "Roberts", "Carter"
    ]

    for i in range(40):
        row = i + 2
        student_id = f"STU-{2024000 + i + 1}"
        name = f"{first_names[i]} {last_names[i]}"
        attendance = round(random.uniform(55, 100), 1)
        hw_avg = round(random.uniform(40, 100), 1)
        test_avg = round(random.uniform(35, 100), 1)
        # Final grade is weighted: 20% attendance, 30% hw, 50% test
        final_grade = round(0.2 * attendance + 0.3 * hw_avg + 0.5 * test_avg, 1)
        at_risk = "Yes" if final_grade < 65 else "No"

        ws.cell(row=row, column=1, value=student_id)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=attendance)
        ws.cell(row=row, column=4, value=hw_avg)
        ws.cell(row=row, column=5, value=test_avg)
        ws.cell(row=row, column=6, value=final_grade)
        ws.cell(row=row, column=7, value=at_risk)

    # Set reasonable column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 13
    ws.column_dimensions["G"].width = 10

    # NO conditional formatting
    # NO class average row
    # NO grade distribution table
    # NO frozen panes

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
