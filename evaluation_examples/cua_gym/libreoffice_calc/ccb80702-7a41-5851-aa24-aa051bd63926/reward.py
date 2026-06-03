"""
Reward Script: Student Performance Tracker Dashboard
Task ID: calc_gsd_041
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Conditional formatting on G2:G41 — Yes=red+white+bold, No=green
  Component 2 (0.25): Class Average row with AVERAGE formulas
  Component 3 (0.25): Grade distribution mini-table with COUNTIF formulas
  Component 4 (0.25): Freeze row 1 (freeze_panes == 'A2')
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_041'


def persist_app_state(domain: str):
    """Best-effort save via Ctrl+S for GUI apps."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Tracker' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Tracker' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Tracker']

    # =========================================================================
    # Component 1: Conditional formatting on G2:G41 (0.25 points)
    # Yes cells -> red bg (FFFF0000) + white bold font; No cells -> green bg
    # =========================================================================
    try:
        cf_list = list(ws.conditional_formatting)
        # Look for CF rules covering G2:G41
        # Track how many valid rules found (0=none, 1=one, 2=both)
        cf_rules_found = 0

        for cf in cf_list:
            cf_range = str(cf)
            # Check if it covers the G column range (G2:G41 or similar)
            if 'G' in cf_range:
                for rule in cf.rules:
                    rule_type = getattr(rule, 'type', '')
                    formula = getattr(rule, 'formula', [])

                    # Check for "Yes" rule
                    if rule_type == 'cellIs' and formula and '"Yes"' in str(formula):
                        dxf = getattr(rule, 'dxf', None)
                        if dxf:
                            # Check red fill
                            if dxf.fill and dxf.fill.fgColor:
                                fg_rgb = str(dxf.fill.fgColor.rgb).upper()
                                if 'FF0000' in fg_rgb:
                                    cf_rules_found += 1
                                    print(f"PASS: 'Yes' CF rule found with red fill ({fg_rgb})")
                                    # Check bold + white font (bonus, not strict)
                                    if dxf.font:
                                        print(f"  Font bold={dxf.font.bold}, color={dxf.font.color.rgb if dxf.font.color else None}")
                                else:
                                    print(f"FAIL: 'Yes' CF rule fill color is {fg_rgb}, expected red")
                        continue

                    # Check for "No" rule
                    if rule_type == 'cellIs' and formula and '"No"' in str(formula):
                        dxf = getattr(rule, 'dxf', None)
                        if dxf:
                            if dxf.fill and dxf.fill.fgColor:
                                fg_rgb = str(dxf.fill.fgColor.rgb).upper()
                                if '00B050' in fg_rgb or '00FF00' in fg_rgb or 'GREEN' in fg_rgb or '008000' in fg_rgb:
                                    cf_rules_found += 1
                                    print(f"PASS: 'No' CF rule found with green fill ({fg_rgb})")
                                else:
                                    print(f"FAIL: 'No' CF rule fill color is {fg_rgb}, expected green")
                        continue

        if cf_rules_found >= 2:
            print(f"PASS: Component 1 — Both CF rules found (0.25 pts)")
            total_score += 0.25
        elif cf_rules_found == 1:
            print(f"PARTIAL: Component 1 — Only one CF rule found (0.125 pts)")
            total_score += 0.125
        else:
            print(f"FAIL: Component 1 — No conditional formatting rules found on G column")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Class Average row with AVERAGE formulas (0.25 points)
    # Expect 'Class Average' label (bold) and AVERAGE formulas in C,D,E,F
    # =========================================================================
    try:
        avg_row = None
        # Search rows 42-50 for the 'Class Average' label
        for r in range(42, 51):
            val = ws.cell(r, 1).value
            if val and 'class average' in str(val).lower():
                avg_row = r
                break

        if avg_row is None:
            print("FAIL: Component 2 — 'Class Average' label not found in rows 42-50")
        else:
            print(f"  Found 'Class Average' in row {avg_row}")
            label_bold = ws.cell(avg_row, 1).font.bold
            if label_bold:
                print(f"  Label is bold: True")

            # Check AVERAGE formulas in columns C, D, E, F (cols 3-6)
            formula_count = 0
            for col_idx in [3, 4, 5, 6]:
                cell_val = ws.cell(avg_row, col_idx).value
                if cell_val and isinstance(cell_val, str) and 'AVERAGE' in cell_val.upper():
                    formula_count += 1
                    print(f"  Col {col_idx} formula: {cell_val}")
                else:
                    print(f"  Col {col_idx}: expected AVERAGE formula, found: {cell_val}")

            if formula_count == 4:
                print(f"PASS: Component 2 — Class Average row with all 4 AVERAGE formulas (0.25 pts)")
                total_score += 0.25
            elif formula_count >= 2:
                partial = 0.25 * (formula_count / 4)
                print(f"PARTIAL: Component 2 — {formula_count}/4 AVERAGE formulas found ({partial:.3f} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 2 — Only {formula_count}/4 AVERAGE formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Grade distribution mini-table with COUNTIF (0.25 points)
    # Rows 45-50: header row + 5 grade bands (A/B/C/D/F) with COUNTIF formulas
    # =========================================================================
    try:
        # Find the header row for grade distribution
        dist_header_row = None
        for r in range(44, 55):
            val = ws.cell(r, 1).value
            if val and 'grade' in str(val).lower() and 'band' in str(val).lower():
                dist_header_row = r
                break
            if val and 'grade' in str(val).lower() and 'dist' in str(val).lower():
                dist_header_row = r
                break

        if dist_header_row is None:
            # Try finding by looking for the pattern: a row with "A" followed by "B", "C", etc.
            for r in range(44, 55):
                val = ws.cell(r, 1).value
                if val and str(val).strip() == 'A':
                    # Check if this starts a grade band sequence
                    next_val = ws.cell(r + 1, 1).value
                    if next_val and str(next_val).strip() == 'B':
                        dist_header_row = r - 1  # header is one row above
                        break

        if dist_header_row is None:
            print("FAIL: Component 3 — Grade distribution header not found")
        else:
            print(f"  Found grade distribution header in row {dist_header_row}")
            # Check header has 'Grade Band' and 'Count' labels
            header_a = ws.cell(dist_header_row, 1).value
            header_b = ws.cell(dist_header_row, 2).value
            print(f"  Header: A={header_a}, B={header_b}")

            # Check grade bands and COUNTIF formulas
            expected_bands = ['A', 'B', 'C', 'D', 'F']
            band_count = 0
            countif_count = 0
            for i, band in enumerate(expected_bands):
                data_row = dist_header_row + 1 + i
                band_val = ws.cell(data_row, 1).value
                formula_val = ws.cell(data_row, 2).value

                if band_val and str(band_val).strip().upper() == band:
                    band_count += 1

                if formula_val and isinstance(formula_val, str) and 'COUNTIF' in formula_val.upper():
                    countif_count += 1
                    print(f"  Band {band} (row {data_row}): formula={formula_val}")
                else:
                    print(f"  Band {band} (row {data_row}): val={band_val}, formula={formula_val}")

            if band_count >= 4 and countif_count >= 4:
                print(f"PASS: Component 3 — Grade distribution table with {band_count} bands, {countif_count} COUNTIF formulas (0.25 pts)")
                total_score += 0.25
            elif countif_count >= 2:
                partial = 0.25 * (countif_count / 5)
                print(f"PARTIAL: Component 3 — {countif_count}/5 COUNTIF formulas ({partial:.3f} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 3 — bands={band_count}, countif={countif_count}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: Freeze row 1 (0.25 points)
    # freeze_panes should be 'A2' (freeze first row)
    # =========================================================================
    try:
        freeze = ws.freeze_panes
        if freeze is not None and str(freeze) == 'A2':
            print(f"PASS: Component 4 — Freeze panes set to A2 (0.25 pts)")
            total_score += 0.25
        elif freeze is not None:
            # Any freeze that freezes row 1 (row component of freeze cell >= 2)
            # A2, B2, etc. all freeze row 1
            from openpyxl.utils import coordinate_to_tuple
            row_idx, _ = coordinate_to_tuple(str(freeze))
            if row_idx == 2:
                print(f"PASS: Component 4 — Freeze panes set to {freeze}, row 1 frozen (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 4 — Freeze panes set to {freeze}, row 1 not frozen")
        else:
            print(f"FAIL: Component 4 — No freeze panes set (found: {freeze})")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
