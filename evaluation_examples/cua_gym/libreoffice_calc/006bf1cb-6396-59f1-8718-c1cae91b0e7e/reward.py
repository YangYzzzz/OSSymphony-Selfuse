"""
Reward Script: Grade Distribution Summary in LibreOffice Calc
Task ID: calc_gsd_022
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20) - Merged header A53:D53 with bold "Grade Distribution Summary"
  Component 2 (0.20) - Grade band labels in A54:A58 (bold)
  Component 3 (0.25) - COUNTIFS formulas in B54:B58
  Component 4 (0.15) - Percentage formulas in C54:C58 with percentage format
  Component 5 (0.20) - Borders applied to A53:C58
"""

import os
import re
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_022'

# Expected grade band labels
EXPECTED_LABELS = [
    'A (90-100)',
    'B (80-89)',
    'C (70-79)',
    'D (60-69)',
    'F (<60)',
]


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Results' not in wb.sheetnames:
        print("CRITICAL: 'Results' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Results']

    # Component 1: Merged header A53:D53 with bold "Grade Distribution Summary" (0.20 points)
    try:
        # Check for merge range covering A53:D53
        merge_found = any(
            mr.min_row == 53 and mr.max_row == 53 and mr.min_col == 1 and mr.max_col >= 4
            for mr in ws.merged_cells.ranges
        )

        header_val = ws['A53'].value
        header_bold = ws['A53'].font.bold

        if merge_found and header_val and 'grade distribution' in str(header_val).lower() and header_bold:
            print(f"PASS: Component 1 — Merged header found: '{header_val}', bold={header_bold} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 — merge={merge_found}, value={repr(header_val)}, bold={header_bold}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Grade band labels in A54:A58, bold (0.20 points)
    try:
        labels_correct = 0
        labels_bold = 0
        for i, expected in enumerate(EXPECTED_LABELS):
            row = 54 + i
            cell = ws.cell(row=row, column=1)
            val = str(cell.value).strip() if cell.value else ''
            # Flexible match: check key parts of each label
            if val.lower() == expected.lower():
                labels_correct += 1
            elif expected.split(' ')[0].lower() in val.lower() and any(c.isdigit() or c == '<' for c in val):
                # Accept reasonable variants like "A: 90-100" or "A (90–100)"
                labels_correct += 1
            if cell.font.bold:
                labels_bold += 1

        if labels_correct == 5 and labels_bold >= 4:
            print(f"PASS: Component 2 — All 5 labels correct ({labels_correct}/5), bold ({labels_bold}/5) (0.20 pts)")
            total_score += 0.20
        elif labels_correct >= 3:
            partial = 0.10
            print(f"PARTIAL: Component 2 — {labels_correct}/5 labels, {labels_bold}/5 bold ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {labels_correct}/5 labels correct, {labels_bold}/5 bold")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: COUNTIFS formulas in B54:B58 (0.25 points)
    try:
        formulas_ok = 0
        for row in range(54, 59):
            cell = ws.cell(row=row, column=2)
            val = str(cell.value).upper().replace(' ', '') if cell.value else ''
            if 'COUNTIF' in val:
                # Verify it references the score column (E)
                if 'E' in val.upper():
                    formulas_ok += 1
                    print(f"  B{row}: COUNTIFS formula found: {cell.value}")
                else:
                    print(f"  B{row}: COUNTIF found but doesn't reference column E: {cell.value}")
            else:
                print(f"  B{row}: No COUNTIF formula, value={repr(cell.value)}")

        if formulas_ok == 5:
            print(f"PASS: Component 3 — All 5 COUNTIFS formulas correct (0.25 pts)")
            total_score += 0.25
        elif formulas_ok >= 3:
            partial = round(0.25 * formulas_ok / 5, 2)
            print(f"PARTIAL: Component 3 — {formulas_ok}/5 formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {formulas_ok}/5 COUNTIFS formulas found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Percentage formulas in C54:C58 with percentage number format (0.15 points)
    try:
        pct_formulas_ok = 0
        pct_format_ok = 0
        for row in range(54, 59):
            cell = ws.cell(row=row, column=3)
            val = str(cell.value).upper().replace(' ', '') if cell.value else ''
            # Check formula references B column and divides by 50 (or uses 50)
            if '=' in val and ('B' in val.upper() or '/' in val):
                pct_formulas_ok += 1
            # Check number format contains '%'
            nf = cell.number_format if cell.number_format else ''
            if '%' in nf:
                pct_format_ok += 1

        if pct_formulas_ok >= 4 and pct_format_ok >= 4:
            print(f"PASS: Component 4 — {pct_formulas_ok}/5 pct formulas, {pct_format_ok}/5 pct format (0.15 pts)")
            total_score += 0.15
        elif pct_formulas_ok >= 3 or pct_format_ok >= 3:
            partial = 0.08
            print(f"PARTIAL: Component 4 — {pct_formulas_ok}/5 formulas, {pct_format_ok}/5 format ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — {pct_formulas_ok}/5 pct formulas, {pct_format_ok}/5 pct format")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Borders applied to A53:C58 (0.20 points)
    try:
        bordered_count = 0
        total_cells = 0
        for row in range(53, 59):
            for col in range(1, 4):
                total_cells += 1
                cell = ws.cell(row=row, column=col)
                b = cell.border
                has_border = any([
                    b.left.style is not None,
                    b.right.style is not None,
                    b.top.style is not None,
                    b.bottom.style is not None,
                ])
                if has_border:
                    bordered_count += 1

        if bordered_count >= 16:  # At least 16/18 cells have some border
            print(f"PASS: Component 5 — {bordered_count}/{total_cells} cells have borders (0.20 pts)")
            total_score += 0.20
        elif bordered_count >= 10:
            partial = 0.10
            print(f"PARTIAL: Component 5 — {bordered_count}/{total_cells} cells bordered ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — Only {bordered_count}/{total_cells} cells have borders")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook for LibreOffice (save unsaved changes)
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
persist_app_state()

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
