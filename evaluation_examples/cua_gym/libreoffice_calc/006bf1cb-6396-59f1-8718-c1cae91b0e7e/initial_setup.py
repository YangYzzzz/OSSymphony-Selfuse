"""
Initial Setup: Create student grade distribution spreadsheet
Task ID: calc_gsd_022
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random
import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_022'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Results'

    # --- Header Row ---
    headers = ['Student ID', 'First Name', 'Last Name', 'Course', 'Overall Grade', 'Letter Grade']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # --- Student Data (50 students, rows 2-51) ---
    first_names = [
        'Sarah', 'Marcus', 'Emily', 'James', 'Olivia', 'Ethan', 'Sophia', 'Liam',
        'Isabella', 'Noah', 'Mia', 'Lucas', 'Charlotte', 'Mason', 'Amelia', 'Logan',
        'Harper', 'Alexander', 'Evelyn', 'Daniel', 'Abigail', 'Matthew', 'Ella',
        'Benjamin', 'Scarlett', 'Henry', 'Grace', 'Sebastian', 'Chloe', 'Jack',
        'Victoria', 'Owen', 'Riley', 'Samuel', 'Aria', 'Ryan', 'Lily', 'Nathan',
        'Zoey', 'Andrew', 'Penelope', 'Gabriel', 'Layla', 'Dylan', 'Nora',
        'Caleb', 'Hannah', 'Isaac', 'Addison', 'Joshua'
    ]
    last_names = [
        'Chen', 'Johnson', 'Williams', 'Patel', 'Garcia', 'Kim', 'Thompson',
        'Martinez', 'Anderson', 'Taylor', 'Thomas', 'Moore', 'Jackson', 'White',
        'Harris', 'Clark', 'Lewis', 'Robinson', 'Walker', 'Young', 'Allen',
        'King', 'Wright', 'Scott', 'Green', 'Baker', 'Adams', 'Nelson', 'Hill',
        'Ramirez', 'Campbell', 'Mitchell', 'Roberts', 'Carter', 'Phillips',
        'Evans', 'Turner', 'Torres', 'Parker', 'Collins', 'Edwards', 'Stewart',
        'Flores', 'Morris', 'Nguyen', 'Murphy', 'Rivera', 'Cook', 'Rogers', 'Morgan'
    ]
    courses = [
        'MATH 201', 'PHYS 101', 'CHEM 301', 'BIO 202', 'CS 150',
        'ENG 110', 'HIST 240', 'ECON 100', 'PSYCH 200', 'STAT 310'
    ]

    # Use a fixed seed for reproducibility
    random.seed(42)

    # Generate scores with a realistic distribution
    scores = []
    for i in range(50):
        # Normal-ish distribution centered around 75
        score = int(random.gauss(75, 12))
        score = max(0, min(100, score))
        scores.append(score)

    def letter_grade(score):
        if score >= 90:
            return 'A'
        elif score >= 80:
            return 'B'
        elif score >= 70:
            return 'C'
        elif score >= 60:
            return 'D'
        else:
            return 'F'

    for i in range(50):
        row = i + 2
        sid = f'STU-{2024000 + i + 1:07d}'
        ws.cell(row=row, column=1, value=sid)
        ws.cell(row=row, column=2, value=first_names[i])
        ws.cell(row=row, column=3, value=last_names[i])
        ws.cell(row=row, column=4, value=random.choice(courses))
        ws.cell(row=row, column=5, value=scores[i])
        ws.cell(row=row, column=6, value=letter_grade(scores[i]))

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 14

    # Rows 52+ are intentionally empty - the task is to create a summary there

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
