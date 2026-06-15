"""
Initial Setup: Conditional formatting priority conflict on spreadsheet
Task ID: calc_tbl_025
Domain: libreoffice_calc

Two conditional formatting rules on B2:B20:
- Rule 1 (higher priority): >50 → yellow fill
- Rule 2 (lower priority): >100 → green fill
Cells above 100 incorrectly show yellow because yellow rule has higher priority.
"""

import os
import shlex
import subprocess
import time

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_025'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Sales Data ---
    ws = wb.active
    ws.title = "Sales Data"

    # Headers
    headers = ["Product", "Revenue ($)", "Units Sold", "Region"]
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows (B column has numeric values for conditional formatting)
    data = [
        ["Wireless Headphones", 245.00, 18, "Northeast"],       # row 2, B2=245
        ["USB-C Hub", 32.50, 45, "West"],                       # row 3, B3=32.50
        ["Mechanical Keyboard", 89.99, 12, "Southeast"],        # row 4, B4=89.99
        ["Monitor Stand", 150.00, 8, "Midwest"],                # row 5, B5=150
        ["Laptop Sleeve", 24.99, 67, "West"],                   # row 6, B6=24.99
        ["Webcam HD", 65.00, 22, "Northeast"],                  # row 7, B7=65
        ["Desk Lamp", 75.00, 31, "Southeast"],                  # row 8, B8=75
        ["Portable SSD", 119.95, 14, "Midwest"],                # row 9, B9=119.95
        ["Mouse Pad XL", 18.50, 88, "West"],                    # row 10, B10=18.50
        ["Bluetooth Speaker", 55.00, 27, "Northeast"],          # row 11, B11=55
        ["Phone Charger", 42.00, 53, "Southeast"],              # row 12, B12=42
        ["HDMI Cable Pack", 15.99, 102, "West"],                # row 13, B13=15.99
        ["Standing Desk Mat", 85.00, 19, "Midwest"],            # row 14, B14=85
        ["Noise Cancelling Earbuds", 199.99, 9, "Northeast"],   # row 15, B15=199.99
        ["Screen Protector Set", 12.50, 145, "Southeast"],      # row 16, B16=12.50
        ["Ergonomic Mouse", 62.00, 34, "West"],                 # row 17, B17=62
        ["Cable Management Kit", 28.75, 41, "Midwest"],         # row 18, B18=28.75
        ["Tablet Stand", 110.00, 16, "Northeast"],              # row 19, B19=110
        ["Power Strip Surge", 47.50, 38, "Southeast"],          # row 20, B20=47.50
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 2:
                cell.number_format = '#,##0.00'

    # Set column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14

    # --- Conditional Formatting ---
    # IMPORTANT: Priority order matters. Lower priority number = higher priority.
    # In the initial (broken) state: yellow (>50) has higher priority than green (>100).
    # So cells >100 show yellow instead of green.

    yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="solid")

    # Rule 1 (higher priority): >50 → yellow
    yellow_rule = CellIsRule(
        operator="greaterThan",
        formula=["50"],
        fill=yellow_fill,
    )
    yellow_rule.priority = 1

    # Rule 2 (lower priority): >100 → green
    green_rule = CellIsRule(
        operator="greaterThan",
        formula=["100"],
        fill=green_fill,
    )
    green_rule.priority = 2

    ws.conditional_formatting.add("B2:B20", yellow_rule)
    ws.conditional_formatting.add("B2:B20", green_rule)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
