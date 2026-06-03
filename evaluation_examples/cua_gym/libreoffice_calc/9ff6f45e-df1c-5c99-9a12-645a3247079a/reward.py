"""
Reward Script: Set left indentation to 0.3cm on cells A1:E20
Task ID: calc_gfl_067
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): All 100 cells in A1:E20 have indent >= 1 (0.3cm equivalent)
  Component 2 (0.3): Indentation is consistent across all 100 cells (same value)
  Component 3 (0.2): Borders remain intact on all cells in A1:E20
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_067'


def persist_app_state(domain: str):
    """Best-effort save of any unsaved GUI edits."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Report' sheet exists
    if 'Report' not in wb.sheetnames:
        print(f"CRITICAL: Sheet 'Report' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Report']

    # Component 1: All 100 cells in A1:E20 have indent >= 1 (0.5 points)
    # In openpyxl, indent=1 corresponds to LibreOffice's 0.3cm left indent.
    # Award proportional credit based on how many cells have the indent set.
    try:
        indented_count = 0
        total_cells = 100  # 20 rows x 5 columns
        for row in range(1, 21):
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                if cell.alignment and cell.alignment.indent is not None and cell.alignment.indent >= 1:
                    indented_count += 1

        ratio = indented_count / total_cells
        if ratio == 1.0:
            print(f"PASS: Component 1 — All {total_cells} cells have indent >= 1 (0.5 pts)")
            total_score += 0.5
        elif ratio >= 0.5:
            partial = round(0.5 * ratio, 2)
            print(f"PARTIAL: Component 1 — {indented_count}/{total_cells} cells indented ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {indented_count}/{total_cells} cells have indent >= 1")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Indentation is consistent across all 100 cells (0.3 points)
    # All cells should have the same indent value, confirming uniform application.
    try:
        indent_values = set()
        for row in range(1, 21):
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                indent_val = cell.alignment.indent if cell.alignment else 0
                indent_values.add(indent_val)

        # Must have exactly one indent value, and it must be > 0 (task-introduced)
        if len(indent_values) == 1 and min(indent_values) >= 1:
            print(f"PASS: Component 2 — All cells have consistent indent={indent_values.pop()} (0.3 pts)")
            total_score += 0.3
        elif len(indent_values) == 1 and min(indent_values) > 0:
            # Indent set but less than expected — partial credit
            print(f"PARTIAL: Component 2 — Consistent indent but value={indent_values.pop()} < 1 (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 — Inconsistent indent values: {indent_values}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Borders remain intact on all cells in A1:E20 (0.2 points)
    # Task requirement: borders must not be affected by the indentation change.
    # Only score this if indent was actually changed (i.e., component 1 passed at least partially).
    # This avoids awarding points to the initial file which already has borders.
    try:
        if indented_count == 0:
            print(f"FAIL: Component 3 — Skipped because no indentation was applied (0 indented cells)")
        else:
            borders_intact = 0
            for row in range(1, 21):
                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col)
                    b = cell.border
                    if (b.left and b.left.style == 'thin' and
                        b.right and b.right.style == 'thin' and
                        b.top and b.top.style == 'thin' and
                        b.bottom and b.bottom.style == 'thin'):
                        borders_intact += 1

            border_ratio = borders_intact / total_cells
            if border_ratio == 1.0:
                print(f"PASS: Component 3 — All {total_cells} cells retain thin borders (0.2 pts)")
                total_score += 0.2
            elif border_ratio >= 0.8:
                partial = round(0.2 * border_ratio, 2)
                print(f"PARTIAL: Component 3 — {borders_intact}/{total_cells} cells have borders ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 3 — Only {borders_intact}/{total_cells} cells have intact borders")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
