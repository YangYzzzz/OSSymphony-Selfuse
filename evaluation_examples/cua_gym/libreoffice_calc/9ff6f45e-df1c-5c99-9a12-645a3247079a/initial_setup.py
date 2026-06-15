"""
Initial Setup: Set cell padding via text indentation in LibreOffice Calc
Task ID: calc_gfl_067
Domain: libreoffice_calc

Creates a spreadsheet with 20 rows x 5 columns of realistic business data,
box borders on A1:E20, and NO indentation (text flush against left border).
"""

import os
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_067'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Report ---
    ws = wb.active
    ws.title = 'Report'

    # Headers
    headers = ['Employee', 'Department', 'Q1 Revenue', 'Q2 Revenue', 'Status']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    # Data rows (19 rows of realistic business data)
    data = [
        ['Sarah Chen', 'Engineering', 45230.50, 51890.75, 'Active'],
        ['Marcus Johnson', 'Marketing', 38750.00, 42100.25, 'Active'],
        ['Elena Rodriguez', 'Sales', 67800.00, 72450.50, 'Active'],
        ['David Kim', 'Engineering', 52100.75, 48920.00, 'On Leave'],
        ['Priya Patel', 'Finance', 41500.00, 43200.00, 'Active'],
        ['James O\'Brien', 'Sales', 58900.25, 61340.00, 'Active'],
        ['Aisha Mohammed', 'Marketing', 35200.00, 37800.50, 'Active'],
        ['Robert Taylor', 'Engineering', 49800.00, 53100.25, 'Active'],
        ['Maria Santos', 'Finance', 44100.50, 46750.00, 'Probation'],
        ['Wei Zhang', 'Sales', 71200.00, 68900.75, 'Active'],
        ['Jennifer Adams', 'Marketing', 36900.25, 39200.00, 'Active'],
        ['Tomasz Kowalski', 'Engineering', 47500.00, 50800.50, 'Active'],
        ['Fatima Al-Hassan', 'Finance', 42300.75, 44900.00, 'Active'],
        ['Michael Brown', 'Sales', 63400.00, 59800.25, 'Active'],
        ['Yuki Tanaka', 'Engineering', 51700.50, 54200.00, 'On Leave'],
        ['Rachel Green', 'Marketing', 37400.00, 40100.75, 'Active'],
        ['Carlos Mendez', 'Finance', 43800.25, 45600.00, 'Active'],
        ['Sophie Laurent', 'Sales', 59100.00, 64300.50, 'Active'],
        ['Ahmed Hassan', 'Engineering', 48200.75, 52400.00, 'Active'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Apply number format for revenue columns (C and D)
    for row in range(2, 21):
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        ws.cell(row=row, column=4).number_format = '#,##0.00'

    # Apply box borders to A1:E20
    thin_side = Side(style='thin', color='000000')
    box_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for row in range(1, 21):
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = box_border

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
