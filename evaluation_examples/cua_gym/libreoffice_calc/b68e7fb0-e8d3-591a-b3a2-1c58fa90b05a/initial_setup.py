"""
Initial Setup: School Safety Incident Log
Task ID: calc_edu_safety_incident_log_057
Domain: libreoffice_calc

Creates an Incidents sheet with 90 rows of realistic school safety incident data.
No summary sheet, no charts, no formulas - just raw incident records.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_safety_incident_log_057'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Incidents ---
    ws = wb.active
    ws.title = 'Incidents'

    # Headers
    headers = ['Incident ID', 'Date', 'Month', 'Location', 'Type', 'Response Time', 'Resolved']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')

    # 90 rows of realistic incident data
    # Types: Injury, Property Damage, Behavioral, Medical, Security
    # Locations: Classroom, Hallway, Cafeteria, Gymnasium, Parking Lot, Bathroom
    # Months: Jan-Oct 2024 (roughly 9 per month for 90 total)
    data = [
        # Jan (9 incidents)
        ['INC-001', '2024-01-08', 'January', 'Gymnasium', 'Injury', 12, 'Yes'],
        ['INC-002', '2024-01-10', 'January', 'Hallway', 'Behavioral', 8, 'Yes'],
        ['INC-003', '2024-01-12', 'January', 'Cafeteria', 'Property Damage', 15, 'Yes'],
        ['INC-004', '2024-01-15', 'January', 'Classroom', 'Medical', 6, 'Yes'],
        ['INC-005', '2024-01-17', 'January', 'Parking Lot', 'Security', 20, 'Yes'],
        ['INC-006', '2024-01-19', 'January', 'Bathroom', 'Behavioral', 10, 'Yes'],
        ['INC-007', '2024-01-22', 'January', 'Gymnasium', 'Injury', 14, 'Yes'],
        ['INC-008', '2024-01-24', 'January', 'Hallway', 'Property Damage', 9, 'Yes'],
        ['INC-009', '2024-01-26', 'January', 'Classroom', 'Behavioral', 7, 'Yes'],
        # Feb (8 incidents)
        ['INC-010', '2024-02-02', 'February', 'Cafeteria', 'Injury', 11, 'Yes'],
        ['INC-011', '2024-02-05', 'February', 'Parking Lot', 'Security', 25, 'No'],
        ['INC-012', '2024-02-07', 'February', 'Hallway', 'Behavioral', 9, 'Yes'],
        ['INC-013', '2024-02-09', 'February', 'Gymnasium', 'Medical', 5, 'Yes'],
        ['INC-014', '2024-02-13', 'February', 'Classroom', 'Property Damage', 18, 'Yes'],
        ['INC-015', '2024-02-15', 'February', 'Bathroom', 'Behavioral', 12, 'Yes'],
        ['INC-016', '2024-02-20', 'February', 'Cafeteria', 'Injury', 8, 'Yes'],
        ['INC-017', '2024-02-23', 'February', 'Hallway', 'Security', 22, 'No'],
        # Mar (10 incidents)
        ['INC-018', '2024-03-01', 'March', 'Gymnasium', 'Injury', 13, 'Yes'],
        ['INC-019', '2024-03-04', 'March', 'Classroom', 'Behavioral', 7, 'Yes'],
        ['INC-020', '2024-03-06', 'March', 'Hallway', 'Property Damage', 16, 'Yes'],
        ['INC-021', '2024-03-08', 'March', 'Cafeteria', 'Medical', 4, 'Yes'],
        ['INC-022', '2024-03-11', 'March', 'Parking Lot', 'Security', 28, 'Yes'],
        ['INC-023', '2024-03-13', 'March', 'Bathroom', 'Behavioral', 10, 'Yes'],
        ['INC-024', '2024-03-15', 'March', 'Gymnasium', 'Injury', 11, 'Yes'],
        ['INC-025', '2024-03-18', 'March', 'Hallway', 'Behavioral', 9, 'Yes'],
        ['INC-026', '2024-03-20', 'March', 'Classroom', 'Medical', 6, 'Yes'],
        ['INC-027', '2024-03-22', 'March', 'Cafeteria', 'Property Damage', 14, 'Yes'],
        # Apr (9 incidents)
        ['INC-028', '2024-04-02', 'April', 'Gymnasium', 'Injury', 10, 'Yes'],
        ['INC-029', '2024-04-04', 'April', 'Hallway', 'Behavioral', 8, 'Yes'],
        ['INC-030', '2024-04-08', 'April', 'Parking Lot', 'Security', 30, 'No'],
        ['INC-031', '2024-04-10', 'April', 'Classroom', 'Medical', 5, 'Yes'],
        ['INC-032', '2024-04-12', 'April', 'Cafeteria', 'Behavioral', 11, 'Yes'],
        ['INC-033', '2024-04-15', 'April', 'Bathroom', 'Property Damage', 17, 'Yes'],
        ['INC-034', '2024-04-17', 'April', 'Gymnasium', 'Injury', 13, 'Yes'],
        ['INC-035', '2024-04-22', 'April', 'Hallway', 'Behavioral', 9, 'Yes'],
        ['INC-036', '2024-04-25', 'April', 'Classroom', 'Security', 24, 'Yes'],
        # May (9 incidents)
        ['INC-037', '2024-05-02', 'May', 'Cafeteria', 'Property Damage', 15, 'Yes'],
        ['INC-038', '2024-05-06', 'May', 'Gymnasium', 'Injury', 12, 'Yes'],
        ['INC-039', '2024-05-08', 'May', 'Hallway', 'Behavioral', 8, 'Yes'],
        ['INC-040', '2024-05-10', 'May', 'Parking Lot', 'Security', 26, 'No'],
        ['INC-041', '2024-05-13', 'May', 'Bathroom', 'Medical', 5, 'Yes'],
        ['INC-042', '2024-05-15', 'May', 'Classroom', 'Behavioral', 10, 'Yes'],
        ['INC-043', '2024-05-17', 'May', 'Gymnasium', 'Injury', 14, 'Yes'],
        ['INC-044', '2024-05-20', 'May', 'Cafeteria', 'Behavioral', 7, 'Yes'],
        ['INC-045', '2024-05-23', 'May', 'Hallway', 'Property Damage', 18, 'Yes'],
        # Jun (7 incidents)
        ['INC-046', '2024-06-03', 'June', 'Gymnasium', 'Injury', 11, 'Yes'],
        ['INC-047', '2024-06-05', 'June', 'Classroom', 'Behavioral', 9, 'Yes'],
        ['INC-048', '2024-06-10', 'June', 'Hallway', 'Security', 21, 'Yes'],
        ['INC-049', '2024-06-12', 'June', 'Cafeteria', 'Medical', 6, 'Yes'],
        ['INC-050', '2024-06-14', 'June', 'Bathroom', 'Property Damage', 13, 'Yes'],
        ['INC-051', '2024-06-18', 'June', 'Parking Lot', 'Security', 29, 'No'],
        ['INC-052', '2024-06-21', 'June', 'Gymnasium', 'Injury', 10, 'Yes'],
        # Jul (6 incidents)
        ['INC-053', '2024-07-02', 'July', 'Parking Lot', 'Security', 32, 'No'],
        ['INC-054', '2024-07-08', 'July', 'Gymnasium', 'Injury', 15, 'Yes'],
        ['INC-055', '2024-07-11', 'July', 'Cafeteria', 'Behavioral', 8, 'Yes'],
        ['INC-056', '2024-07-15', 'July', 'Hallway', 'Property Damage', 12, 'Yes'],
        ['INC-057', '2024-07-18', 'July', 'Classroom', 'Medical', 5, 'Yes'],
        ['INC-058', '2024-07-22', 'July', 'Bathroom', 'Behavioral', 11, 'Yes'],
        # Aug (8 incidents)
        ['INC-059', '2024-08-05', 'August', 'Gymnasium', 'Injury', 13, 'Yes'],
        ['INC-060', '2024-08-07', 'August', 'Hallway', 'Behavioral', 9, 'Yes'],
        ['INC-061', '2024-08-09', 'August', 'Classroom', 'Property Damage', 16, 'Yes'],
        ['INC-062', '2024-08-12', 'August', 'Cafeteria', 'Medical', 4, 'Yes'],
        ['INC-063', '2024-08-14', 'August', 'Parking Lot', 'Security', 27, 'No'],
        ['INC-064', '2024-08-19', 'August', 'Bathroom', 'Behavioral', 10, 'Yes'],
        ['INC-065', '2024-08-21', 'August', 'Gymnasium', 'Injury', 12, 'Yes'],
        ['INC-066', '2024-08-26', 'August', 'Hallway', 'Security', 23, 'Yes'],
        # Sep (11 incidents - the most)
        ['INC-067', '2024-09-02', 'September', 'Classroom', 'Behavioral', 8, 'Yes'],
        ['INC-068', '2024-09-04', 'September', 'Gymnasium', 'Injury', 11, 'Yes'],
        ['INC-069', '2024-09-06', 'September', 'Hallway', 'Property Damage', 14, 'Yes'],
        ['INC-070', '2024-09-09', 'September', 'Cafeteria', 'Medical', 5, 'Yes'],
        ['INC-071', '2024-09-11', 'September', 'Parking Lot', 'Security', 31, 'No'],
        ['INC-072', '2024-09-13', 'September', 'Bathroom', 'Behavioral', 12, 'Yes'],
        ['INC-073', '2024-09-16', 'September', 'Gymnasium', 'Injury', 10, 'Yes'],
        ['INC-074', '2024-09-18', 'September', 'Hallway', 'Behavioral', 7, 'Yes'],
        ['INC-075', '2024-09-20', 'September', 'Classroom', 'Security', 25, 'Yes'],
        ['INC-076', '2024-09-23', 'September', 'Cafeteria', 'Property Damage', 16, 'Yes'],
        ['INC-077', '2024-09-25', 'September', 'Gymnasium', 'Injury', 9, 'Yes'],
        # Oct (7 incidents)
        ['INC-078', '2024-10-02', 'October', 'Classroom', 'Behavioral', 10, 'Yes'],
        ['INC-079', '2024-10-07', 'October', 'Hallway', 'Injury', 12, 'Yes'],
        ['INC-080', '2024-10-09', 'October', 'Gymnasium', 'Medical', 6, 'Yes'],
        ['INC-081', '2024-10-11', 'October', 'Cafeteria', 'Property Damage', 15, 'Yes'],
        ['INC-082', '2024-10-14', 'October', 'Parking Lot', 'Security', 28, 'No'],
        ['INC-083', '2024-10-16', 'October', 'Bathroom', 'Behavioral', 9, 'Yes'],
        ['INC-084', '2024-10-21', 'October', 'Gymnasium', 'Injury', 13, 'Yes'],
        # Nov (6 incidents)
        ['INC-085', '2024-11-04', 'November', 'Classroom', 'Behavioral', 8, 'Yes'],
        ['INC-086', '2024-11-06', 'November', 'Hallway', 'Property Damage', 17, 'Yes'],
        ['INC-087', '2024-11-08', 'November', 'Gymnasium', 'Injury', 11, 'Yes'],
        ['INC-088', '2024-11-12', 'November', 'Cafeteria', 'Medical', 5, 'Yes'],
        ['INC-089', '2024-11-15', 'November', 'Parking Lot', 'Security', 24, 'No'],
        ['INC-090', '2024-11-19', 'November', 'Bathroom', 'Behavioral', 10, 'Yes'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 10

    # Freeze the header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Incidents sheet: 90 rows of incident data')


create_initial()
