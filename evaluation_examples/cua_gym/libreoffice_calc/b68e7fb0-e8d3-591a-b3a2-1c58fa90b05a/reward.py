"""
Reward Script: Analyze safety incident reports and create summary sheet with charts
Task ID: calc_edu_safety_incident_log_057
Domain: libreoffice_calc
Scoring:
  - Component 1: 'Incident Summary' sheet exists (0.25 pts)
  - Component 2: Count by Type table with COUNTIF formulas for all 5 types (0.25 pts)
  - Component 3: Average response time using AVERAGE(Incidents!F2:F91) formula (0.20 pts)
  - Component 4: Month with most incidents using INDEX/MATCH/MAX formula (0.15 pts)
  - Component 5: Bar chart on Summary sheet with title 'Safety Incidents by Type' (0.15 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_safety_incident_log_057'


def find_formula_in_sheet(ws, keyword_list):
    """
    Scan all cells in ws for a formula containing ALL keywords (case-insensitive, space-stripped).
    Returns the cell value if found, else None.
    """
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val_upper = cell.value.upper().replace(' ', '')
                if all(kw.upper().replace(' ', '') in val_upper for kw in keyword_list):
                    return cell.value
    return None


def count_formulas_in_sheet(ws, keyword_list):
    """
    Count cells in ws whose formula contains ALL keywords (case-insensitive, space-stripped).
    """
    count = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val_upper = cell.value.upper().replace(' ', '')
                if all(kw.upper().replace(' ', '') in val_upper for kw in keyword_list):
                    count += 1
    return count


def extract_chart_title(chart):
    """
    Attempt to extract chart title text. Returns None if not accessible.
    """
    try:
        return chart.title.tx.rich.p[0].r[0].t
    except Exception:
        return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: 'Incident Summary' sheet exists (0.25 points)
    # The initial file only has 'Incidents' sheet. Task requires creating 'Incident Summary' sheet.
    try:
        if 'Incident Summary' in wb.sheetnames:
            print("PASS: Component 1 — 'Incident Summary' sheet exists (0.25 pts)")
            total_score += 0.25
            ws_summary = wb['Incident Summary']
        else:
            print(f"FAIL: Component 1 — 'Incident Summary' sheet not found. Sheets: {wb.sheetnames}")
            # No summary sheet means remaining components cannot pass
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Component 2: Count by Type table with COUNTIF formulas (0.25 points)
    # Must have COUNTIF formulas referencing Incidents!E column for all 5 incident types.
    # The golden file shows B3:B7 all use =COUNTIF(Incidents!E2:E91,A<row>).
    try:
        countif_type_count = count_formulas_in_sheet(ws_summary, ['COUNTIF', 'INCIDENTS!E'])
        expected_types = 5  # Injury, Property Damage, Behavioral, Medical, Security

        if countif_type_count >= expected_types:
            print(f"PASS: Component 2 — Count by Type: found {countif_type_count} COUNTIF formulas referencing Incidents!E (0.25 pts)")
            total_score += 0.25
        elif countif_type_count > 0:
            partial = round(0.25 * countif_type_count / expected_types, 4)
            print(f"PARTIAL: Component 2 — Only {countif_type_count}/{expected_types} COUNTIF formulas for types found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No COUNTIF formulas referencing Incidents!E column found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Average response time formula (0.20 points)
    # Must have an AVERAGE formula referencing Incidents!F column (specifically F2:F91).
    try:
        avg_formula = find_formula_in_sheet(ws_summary, ['AVERAGE', 'INCIDENTS!', '!F'])
        if avg_formula is not None:
            print(f"PASS: Component 3 — Average response time formula found: {avg_formula} (0.20 pts)")
            total_score += 0.20
        else:
            print("FAIL: Component 3 — No AVERAGE formula referencing Incidents!F column found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Month with most incidents using INDEX/MATCH/MAX formula (0.15 points)
    # Must have a formula that identifies the month with the highest incident count.
    # Acceptable: INDEX+MATCH+MAX, or VLOOKUP+MAX with COUNTIF-by-month data.
    try:
        index_match_formula = find_formula_in_sheet(ws_summary, ['INDEX', 'MATCH', 'MAX'])

        if index_match_formula is not None:
            print(f"PASS: Component 4 — Month with most incidents formula found: {index_match_formula} (0.15 pts)")
            total_score += 0.15
        else:
            # Check if there are COUNTIF per month (referencing C column) + a separate MAX formula
            countif_month_count = count_formulas_in_sheet(ws_summary, ['COUNTIF', 'INCIDENTS!C'])
            max_formula = find_formula_in_sheet(ws_summary, ['MAX'])

            if countif_month_count >= 6 and max_formula is not None:
                print(f"PARTIAL: Component 4 — {countif_month_count} month COUNTIFs + MAX formula present, but no INDEX/MATCH result cell (0.08 pts)")
                total_score += 0.08
            elif countif_month_count >= 6:
                print(f"PARTIAL: Component 4 — {countif_month_count} month COUNTIF formulas present but no MAX/result formula (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 4 — No month analysis formula found (INDEX/MATCH/MAX or COUNTIF by month)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Bar chart on Summary sheet with title 'Safety Incidents by Type' (0.15 points)
    try:
        charts = ws_summary._charts
        if len(charts) == 0:
            print("FAIL: Component 5 — No charts found on 'Incident Summary' sheet")
        else:
            # Find a bar chart
            bar_charts = [c for c in charts if 'Bar' in c.__class__.__name__]
            other_charts = [c for c in charts if 'Bar' not in c.__class__.__name__]

            if len(bar_charts) > 0:
                chart = bar_charts[0]
                title_text = extract_chart_title(chart)

                if title_text is not None and 'Safety Incidents by Type' in title_text:
                    print(f"PASS: Component 5 — Bar chart with title '{title_text}' found on Summary sheet (0.15 pts)")
                    total_score += 0.15
                elif title_text is not None and 'Safety Incidents by Type' not in title_text:
                    print(f"PARTIAL: Component 5 — Bar chart found but title is '{title_text}', expected 'Safety Incidents by Type' (0.08 pts)")
                    total_score += 0.08
                elif title_text is None:
                    print(f"PARTIAL: Component 5 — Bar chart found but title could not be extracted (0.08 pts)")
                    total_score += 0.08
            elif len(other_charts) > 0:
                chart_types = [c.__class__.__name__ for c in other_charts]
                print(f"FAIL: Component 5 — Chart(s) found but none are BarChart: {chart_types}")
            else:
                print("FAIL: Component 5 — No charts found on 'Incident Summary' sheet")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
