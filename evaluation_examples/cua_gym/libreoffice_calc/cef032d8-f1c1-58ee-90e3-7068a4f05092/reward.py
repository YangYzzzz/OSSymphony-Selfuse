"""
Reward Script: Pivot Table (DataPilot) creation from Transactions data
Task ID: calc_ggf_024
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20): A new pivot-table sheet exists beyond 'Transactions'
  Component 2 (0.30): Correct structure — 5 category rows x 12 month columns with headers
  Component 3 (0.30): Summed Amount values are correct (spot-check multiple cells)
  Component 4 (0.20): Grand Total row and/or column present with correct values
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_024'

# Expected categories and month order
EXPECTED_CATEGORIES = {'Clothing', 'Electronics', 'Food', 'Home', 'Sports'}
EXPECTED_MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                   'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

# Ground truth pivot values (from Transactions data aggregation)
# Format: (category, month) -> expected_sum
SPOT_CHECK_VALUES = {
    ('Electronics', 'Jan'): 1669.82,
    ('Clothing', 'Feb'): 238.0,
    ('Food', 'Mar'): 44.33,
    ('Home', 'Sep'): 1069.71,
    ('Sports', 'Jul'): 791.10,
    ('Electronics', 'Sep'): 2448.28,
}

# Grand total for the entire table
EXPECTED_GRAND_TOTAL = 25586.38


def compute_pivot_from_source(wb):
    """Compute the expected pivot table from source Transactions data."""
    ws = wb['Transactions']
    pivot = {}
    for r in range(2, ws.max_row + 1):
        month = ws.cell(row=r, column=2).value
        cat = ws.cell(row=r, column=3).value
        amt = ws.cell(row=r, column=4).value
        if cat and month and amt is not None:
            key = (str(cat).strip(), str(month).strip())
            pivot[key] = pivot.get(key, 0.0) + float(amt)
    return pivot


def find_pivot_sheet(wb):
    """Find the pivot table sheet (any sheet that is not 'Transactions')."""
    for name in wb.sheetnames:
        if name.lower() != 'transactions':
            return wb[name]
    return None


def parse_pivot_sheet(ws):
    """Parse the pivot sheet to extract category-month mapping.
    Returns: dict of {(category, month): value}, set of categories found,
             list of months found, and grand_total if present.
    """
    # Row 1 should be headers: first cell is category label, rest are months
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        headers.append(str(v).strip() if v is not None else '')

    # Identify month columns
    month_cols = {}  # month_name -> column_index (0-based in headers)
    for i, h in enumerate(headers):
        if h in EXPECTED_MONTHS:
            month_cols[h] = i + 1  # 1-based column

    # Find grand total column
    grand_total_col = None
    for i, h in enumerate(headers):
        if 'grand' in h.lower() and 'total' in h.lower():
            grand_total_col = i + 1

    # Parse data rows
    data = {}
    categories_found = set()
    grand_total_row_value = None

    for r in range(2, ws.max_row + 1):
        row_label = ws.cell(row=r, column=1).value
        if row_label is None:
            continue
        row_label_str = str(row_label).strip()

        if 'grand' in row_label_str.lower() and 'total' in row_label_str.lower():
            # This is the grand total row
            if grand_total_col:
                gt_val = ws.cell(row=r, column=grand_total_col).value
                if gt_val is not None:
                    try:
                        grand_total_row_value = float(gt_val)
                    except (ValueError, TypeError):
                        pass
            continue

        if row_label_str in EXPECTED_CATEGORIES:
            categories_found.add(row_label_str)
            for month_name, col_idx in month_cols.items():
                val = ws.cell(row=r, column=col_idx).value
                if val is not None:
                    try:
                        data[(row_label_str, month_name)] = float(val)
                    except (ValueError, TypeError):
                        pass

    return data, categories_found, list(month_cols.keys()), grand_total_col is not None, grand_total_row_value


def verify_task(file_path):
    """Verify pivot table creation with progressive scoring."""
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that Transactions sheet still exists (precondition gate)
    if 'Transactions' not in wb.sheetnames:
        print("CRITICAL: 'Transactions' sheet missing — file is corrupted")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: A new pivot-table sheet exists (0.2 points)
    try:
        pivot_ws = find_pivot_sheet(wb)
        if pivot_ws is not None:
            print(f"PASS: Component 1 — New sheet found: '{pivot_ws.title}' (0.2 pts)")
            total_score += 0.2
        else:
            print("FAIL: Component 1 — No sheet besides 'Transactions' found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # If no pivot sheet, remaining checks can't proceed
    if pivot_ws is None:
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    # Parse the pivot sheet
    try:
        pivot_data, cats_found, months_found, has_gt_col, gt_row_val = parse_pivot_sheet(pivot_ws)
    except Exception as e:
        print(f"ERROR: Could not parse pivot sheet: {e}")
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    # Component 2: Correct structure — 5 categories x 12 months (0.3 points)
    try:
        cats_ok = cats_found == EXPECTED_CATEGORIES
        months_ok = len(months_found) == 12 and set(months_found) == set(EXPECTED_MONTHS)

        sub_score = 0.0
        if cats_ok:
            sub_score += 0.15
            print(f"PASS: Component 2a — All 5 categories found: {sorted(cats_found)} (0.15 pts)")
        else:
            print(f"FAIL: Component 2a — Expected categories {sorted(EXPECTED_CATEGORIES)}, found {sorted(cats_found)}")

        if months_ok:
            sub_score += 0.15
            print(f"PASS: Component 2b — All 12 months found as columns (0.15 pts)")
        else:
            print(f"FAIL: Component 2b — Expected 12 months, found {len(months_found)}: {months_found}")

        if sub_score > 0:
            total_score += sub_score
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Summed Amount values correct — spot-check (0.3 points)
    try:
        # Also compute expected values from source data for full verification
        expected_pivot = compute_pivot_from_source(wb)

        checks_passed = 0
        checks_total = len(SPOT_CHECK_VALUES)
        tolerance = 0.05

        for (cat, month), expected_val in SPOT_CHECK_VALUES.items():
            actual_val = pivot_data.get((cat, month))
            if actual_val is not None and abs(actual_val - expected_val) <= tolerance:
                checks_passed += 1
                print(f"  PASS: {cat}/{month} = {actual_val} (expected {expected_val})")
            else:
                print(f"  FAIL: {cat}/{month} = {actual_val} (expected {expected_val})")

        if checks_total > 0:
            ratio = checks_passed / checks_total
            pts = round(0.3 * ratio, 2)
            if pts > 0:
                print(f"PASS: Component 3 — {checks_passed}/{checks_total} spot checks passed ({pts} pts)")
                total_score += pts
            else:
                print(f"FAIL: Component 3 — 0/{checks_total} spot checks passed")
        else:
            print("FAIL: Component 3 — No spot checks defined")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Grand Total row/column present with correct values (0.2 points)
    try:
        if has_gt_col and gt_row_val is not None:
            if abs(gt_row_val - EXPECTED_GRAND_TOTAL) <= 0.1:
                print(f"PASS: Component 4 — Grand Total = {gt_row_val} (expected {EXPECTED_GRAND_TOTAL}) (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 4 — Grand Total = {gt_row_val}, expected {EXPECTED_GRAND_TOTAL}")
        elif has_gt_col:
            # Grand total column exists but no grand total row value found
            # Check if row grand totals exist for categories instead
            print(f"FAIL: Component 4 — Grand Total column exists but Grand Total row value not found")
        else:
            print("FAIL: Component 4 — No Grand Total column/row found in pivot table")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point — persist app state first, then verify
def persist_app_state(domain):
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
