"""
Initial Setup: Create store transactions spreadsheet for pivot table task
Task ID: calc_ggf_024
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_024'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

CATEGORIES = ['Electronics', 'Clothing', 'Food', 'Home', 'Sports']
MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
          'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def generate_transactions():
    """Generate 200 realistic transaction rows with fixed seed."""
    random.seed(42)
    rows = []
    for i in range(1, 201):
        txn_id = f'TXN-{i:04d}'
        month = random.choice(MONTHS)
        category = random.choice(CATEGORIES)
        # Amount varies by category for realism
        base = {
            'Electronics': (80, 500),
            'Clothing': (20, 150),
            'Food': (5, 60),
            'Home': (30, 300),
            'Sports': (15, 200),
        }[category]
        amount = round(random.uniform(base[0], base[1]), 2)
        rows.append([txn_id, month, category, amount])
    return rows


def create_initial():
    wb = openpyxl.Workbook()

    # --- Transactions sheet ---
    ws = wb.active
    ws.title = 'Transactions'

    # Header row
    headers = ['TransactionID', 'Month', 'Category', 'Amount']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Data rows (200 rows, rows 2-201)
    data = generate_transactions()
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
