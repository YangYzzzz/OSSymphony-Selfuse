"""
Initial Setup: HR Monthly Turnover Rate Spreadsheet
Task ID: calc_hr_turnover_trend_chart_018
Domain: libreoffice_calc

Creates a Turnover sheet with Month, Separations, Avg Headcount, and Turnover Rate %
columns. D2:D13 (Turnover Rate %) is left EMPTY — the agent must fill in formulas.
No chart exists in the initial file.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_turnover_trend_chart_018'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Turnover ---
    ws = wb.active
    ws.title = 'Turnover'

    # Headers
    headers = ['Month', 'Separations', 'Avg Headcount', 'Turnover Rate %']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Monthly data — realistic HR numbers for a mid-size company (~250 employees)
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    separations = [6, 4, 7, 9, 11, 8, 10, 12, 7, 5, 6, 9]
    avg_headcount = [248, 245, 247, 252, 258, 254, 260, 265, 261, 257, 254, 256]

    for i, (month, seps, headcount) in enumerate(zip(months, separations, avg_headcount), 2):
        ws.cell(row=i, column=1, value=month)
        ws.cell(row=i, column=2, value=seps)
        ws.cell(row=i, column=3, value=headcount)
        # Column D (Turnover Rate %) is intentionally left empty

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    # Freeze the header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
