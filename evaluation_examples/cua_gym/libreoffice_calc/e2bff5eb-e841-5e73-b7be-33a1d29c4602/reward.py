"""
Reward Script: Create a line chart showing monthly turnover rate for the past 12 months
Task ID: calc_hr_turnover_trend_chart_018
Domain: libreoffice_calc

Scoring rubric:
  Component 1: D2:D13 contain formulas =(Bx/Cx)*100  — 0.35 points
  Component 2: D2:D13 are formatted to 1 decimal place — 0.15 points
  Component 3: A line chart exists on the Turnover sheet — 0.20 points
  Component 4: Chart title is 'Monthly Turnover Rate'   — 0.10 points
  Component 5: X-axis label='Month', Y-axis label='Turnover Rate (%)'  — 0.20 points
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_turnover_trend_chart_018'


def get_title_text(title_obj):
    """Extract plain text from an openpyxl chart Title object."""
    try:
        return title_obj.tx.rich.p[0].r[0].t
    except Exception:
        pass
    # Fallback: try strRef
    try:
        return title_obj.tx.strRef.v
    except Exception:
        pass
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Turnover' sheet must exist
    if 'Turnover' not in wb.sheetnames:
        print("FAIL: 'Turnover' sheet not found in workbook")
        print("\nScore: 0.0/1.0")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Turnover']

    # Component 1: D2:D13 contain formulas =(Bx/Cx)*100 (0.35 points)
    # Each row formula should match =(B{row}/C{row})*100
    try:
        formula_pattern = re.compile(
            r'^\s*=\s*\(\s*B(\d+)\s*/\s*C(\d+)\s*\)\s*\*\s*100\s*$',
            re.IGNORECASE
        )
        correct_formulas = 0
        formula_results = []
        for row in range(2, 14):
            val = ws.cell(row=row, column=4).value
            if isinstance(val, str):
                m = formula_pattern.match(val)
                if m and m.group(1) == str(row) and m.group(2) == str(row):
                    correct_formulas += 1
                    formula_results.append(f"D{row}: {val} OK")
                else:
                    formula_results.append(f"D{row}: {repr(val)} WRONG pattern")
            else:
                formula_results.append(f"D{row}: {repr(val)} NOT a formula")

        if correct_formulas == 12:
            print(f"PASS: Component 1 — All 12 formulas =(Bx/Cx)*100 present in D2:D13 (0.35 pts)")
            total_score += 0.35
        elif correct_formulas >= 6:
            partial = round(0.35 * correct_formulas / 12, 4)
            print(f"PARTIAL: Component 1 — {correct_formulas}/12 correct formulas (partial {partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {correct_formulas}/12 formulas correct in D2:D13")
            for r in formula_results:
                print(f"  {r}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: D2:D13 number format is '0.0' (1 decimal place) (0.15 points)
    try:
        correct_formats = 0
        for row in range(2, 14):
            cell_fmt = ws.cell(row=row, column=4).number_format
            # Accept '0.0' or similar 1-decimal formats
            if cell_fmt in ('0.0', '0.0;', '#,##0.0'):
                correct_formats += 1

        if correct_formats == 12:
            print(f"PASS: Component 2 — All 12 cells D2:D13 formatted to 1 decimal place (0.15 pts)")
            total_score += 0.15
        elif correct_formats > 0:
            partial = round(0.15 * correct_formats / 12, 4)
            print(f"PARTIAL: Component 2 — {correct_formats}/12 cells have 1-decimal format (partial {partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — D2:D13 not formatted to 1 decimal place")
            sample_fmt = ws.cell(row=2, column=4).number_format
            print(f"  Sample D2 format: {repr(sample_fmt)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: A line chart exists on the Turnover sheet (0.20 points)
    try:
        charts = ws._charts
        line_charts = [c for c in charts if type(c).__name__ == 'LineChart']

        if len(line_charts) >= 1:
            print(f"PASS: Component 3 — Line chart found on Turnover sheet (0.20 pts)")
            total_score += 0.20
        elif len(charts) >= 1:
            print(f"FAIL: Component 3 — Chart found but not a LineChart (found: {[type(c).__name__ for c in charts]})")
        else:
            print(f"FAIL: Component 3 — No chart found on Turnover sheet")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Chart title is 'Monthly Turnover Rate' (0.10 points)
    try:
        charts = ws._charts
        line_charts = [c for c in charts if type(c).__name__ == 'LineChart']
        if line_charts:
            chart = line_charts[0]
            title_text = get_title_text(chart.title)
            if title_text is not None and title_text.strip() == 'Monthly Turnover Rate':
                print(f"PASS: Component 4 — Chart title is 'Monthly Turnover Rate' (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 4 — Chart title expected 'Monthly Turnover Rate', found: {repr(title_text)}")
        else:
            print(f"SKIP: Component 4 — No line chart to check title on")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: X-axis label = 'Month', Y-axis label = 'Turnover Rate (%)' (0.20 points)
    try:
        charts = ws._charts
        line_charts = [c for c in charts if type(c).__name__ == 'LineChart']
        if line_charts:
            chart = line_charts[0]
            x_title = get_title_text(chart.x_axis.title) if chart.x_axis.title else None
            y_title = get_title_text(chart.y_axis.title) if chart.y_axis.title else None

            x_ok = x_title is not None and x_title.strip() == 'Month'
            y_ok = y_title is not None and y_title.strip() == 'Turnover Rate (%)'

            if x_ok and y_ok:
                print(f"PASS: Component 5 — X-axis='Month', Y-axis='Turnover Rate (%)' (0.20 pts)")
                total_score += 0.20
            elif x_ok:
                print(f"PARTIAL: Component 5 — X-axis='Month' OK but Y-axis wrong: {repr(y_title)} (0.10 pts)")
                total_score += 0.10
            elif y_ok:
                print(f"PARTIAL: Component 5 — Y-axis='Turnover Rate (%)' OK but X-axis wrong: {repr(x_title)} (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 5 — X-axis={repr(x_title)}, Y-axis={repr(y_title)}")
        else:
            print(f"SKIP: Component 5 — No line chart to check axis labels on")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
