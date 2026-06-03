"""
Reward Script: Conditional formatting rule on Comparison sheet
Task ID: calc_gg2_049
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): CF rule exists on range B2:B52
  Component 2 (0.3): CF formula matches ABS(B2-C2)/IF(C2=0,1,C2)>0.1
  Component 3 (0.2): CF style has yellow background fill
  Component 4 (0.2): CF style has bold font
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg2_049'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Comparison' sheet must exist
    if 'Comparison' not in wb.sheetnames:
        print(f"FAIL: 'Comparison' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Comparison']

    # Collect all conditional formatting rules
    cf_list = list(ws.conditional_formatting)
    print(f"INFO: Found {len(cf_list)} conditional formatting rule(s)")

    # Find the target CF rule that covers B2:B52
    target_cf = None
    target_rule = None
    for cf in cf_list:
        cf_range_str = str(cf).strip()
        # Check if the range covers B2:B52 (exact or containing)
        for rule in cf.rules:
            if rule.type == 'expression' and rule.formula:
                # This is a formula-based CF rule
                # Check if range includes B2:B52
                if 'B2:B52' in cf_range_str or 'B2:B52' == cf_range_str.replace('<ConditionalFormatting ', '').rstrip('>'):
                    target_cf = cf
                    target_rule = rule
                    break
        if target_rule:
            break

    # Component 1: CF rule exists on range B2:B52 (0.3 points)
    try:
        if target_cf is not None and target_rule is not None:
            print(f"PASS: Component 1 - CF rule found on range B2:B52 (0.3 pts)")
            total_score += 0.3
        else:
            # Also check if there's any formula-based CF on a range that includes B2:B52
            found_any = False
            for cf in cf_list:
                for rule in cf.rules:
                    if rule.type == 'expression':
                        print(f"  INFO: Found expression CF on range: {cf}, formula: {rule.formula}")
                        found_any = True
            if not found_any:
                print(f"FAIL: Component 1 - No formula-based conditional formatting found on B2:B52")
            else:
                print(f"FAIL: Component 1 - Formula-based CF exists but not on range B2:B52")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: CF formula matches expected expression (0.3 points)
    try:
        if target_rule is not None and target_rule.formula:
            formula_str = str(target_rule.formula[0]).strip()
            # Normalize: remove spaces, uppercase
            normalized = formula_str.upper().replace(' ', '')
            expected = 'ABS(B2-C2)/IF(C2=0,1,C2)>0.1'
            expected_norm = expected.upper().replace(' ', '')

            if normalized == expected_norm:
                print(f"PASS: Component 2 - Formula matches exactly: {formula_str} (0.3 pts)")
                total_score += 0.3
            else:
                # Check for close variants (e.g., with = prefix or different spacing)
                # Strip leading = if present
                cleaned = normalized.lstrip('=')
                if cleaned == expected_norm:
                    print(f"PASS: Component 2 - Formula matches (with leading =): {formula_str} (0.3 pts)")
                    total_score += 0.3
                else:
                    print(f"FAIL: Component 2 - Formula mismatch. Expected: {expected}, Found: {formula_str}")
        else:
            print(f"FAIL: Component 2 - No formula rule found to check")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: CF style has yellow background fill (0.2 points)
    try:
        if target_rule is not None and hasattr(target_rule, 'dxf') and target_rule.dxf:
            dxf = target_rule.dxf
            if dxf.fill and dxf.fill.fgColor:
                fg_rgb = dxf.fill.fgColor.rgb
                pattern_type = dxf.fill.patternType
                # Yellow is FFFFFF00 in ARGB
                # Accept various yellow shades
                is_yellow = False
                if fg_rgb:
                    rgb_str = str(fg_rgb).upper()
                    # FFFFFF00 = pure yellow, also accept FFFFE599, FFFFFFCC etc
                    if rgb_str in ('FFFFFF00', '00FFFF00'):
                        is_yellow = True
                    elif rgb_str.endswith('FFFF00'):
                        is_yellow = True

                if is_yellow and pattern_type == 'solid':
                    print(f"PASS: Component 3 - Yellow fill detected (rgb={fg_rgb}, pattern={pattern_type}) (0.2 pts)")
                    total_score += 0.2
                else:
                    print(f"FAIL: Component 3 - Fill not yellow/solid. rgb={fg_rgb}, pattern={pattern_type}")
            else:
                print(f"FAIL: Component 3 - No fill or fgColor in differential style")
        else:
            print(f"FAIL: Component 3 - No differential style found on the rule")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: CF style has bold font (0.2 points)
    try:
        if target_rule is not None and hasattr(target_rule, 'dxf') and target_rule.dxf:
            dxf = target_rule.dxf
            if dxf.font and dxf.font.bold:
                print(f"PASS: Component 4 - Bold font detected in CF style (0.2 pts)")
                total_score += 0.2
            else:
                font_bold = dxf.font.bold if dxf.font else None
                print(f"FAIL: Component 4 - Font bold not set. font.bold={font_bold}")
        else:
            print(f"FAIL: Component 4 - No differential style found on the rule")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist unsaved state before verification
persist_app_state("libreoffice_calc")

# Run verification
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
