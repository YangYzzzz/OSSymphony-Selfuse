"""
Initial Setup: Conditional formatting for variance comparison
Task ID: calc_gg2_049
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg2_049'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Comparison ---
    ws = wb.active
    ws.title = 'Comparison'

    # Headers
    headers = ['Label', 'System A Value', 'System B Value']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 51 rows of realistic comparison data (rows 2-52)
    # Mix of values: some within 10% tolerance, some outside, include a zero case
    data = [
        ['CPU Utilization (%)', 72.5, 70.1],
        ['Memory Usage (GB)', 15.8, 14.2],
        ['Disk Read IOPS', 4500, 4480],
        ['Disk Write IOPS', 3200, 2800],
        ['Network In (Mbps)', 850, 840],
        ['Network Out (Mbps)', 620, 550],
        ['Request Latency (ms)', 45.3, 44.8],
        ['Error Rate (%)', 0.12, 0.11],
        ['Throughput (req/s)', 12500, 12480],
        ['Cache Hit Ratio (%)', 94.7, 85.0],
        ['Active Connections', 340, 338],
        ['Queue Depth', 28, 25],
        ['Thread Pool Size', 64, 64],
        ['GC Pause (ms)', 120, 95],
        ['Heap Usage (MB)', 2048, 2040],
        ['API Response Time (ms)', 230, 210],
        ['DB Query Time (ms)', 15.4, 14.9],
        ['Session Count', 1580, 1575],
        ['Page Load Time (s)', 3.2, 2.8],
        ['Bounce Rate (%)', 42.1, 41.5],
        ['Conversion Rate (%)', 2.8, 2.5],
        ['Cart Abandonment (%)', 68.5, 62.0],
        ['Avg Order Value ($)', 127.50, 125.00],
        ['Revenue per Visit ($)', 3.57, 3.12],
        ['Customer Satisfaction', 4.2, 4.1],
        ['NPS Score', 45, 42],
        ['Support Tickets', 230, 228],
        ['Resolution Time (hrs)', 4.5, 3.8],
        ['Uptime (%)', 99.95, 99.92],
        ['Deployment Frequency', 12, 11],
        ['Lead Time (days)', 3.5, 3.4],
        ['Change Failure Rate (%)', 8.2, 7.5],
        ['MTTR (minutes)', 35, 30],
        ['Code Coverage (%)', 78.5, 78.0],
        ['Technical Debt (hrs)', 340, 310],
        ['Sprint Velocity', 42, 40],
        ['Defect Density', 0.8, 0.75],
        ['Test Pass Rate (%)', 96.2, 96.0],
        ['Build Time (min)', 8.5, 8.3],
        ['Container CPU (%)', 55.0, 50.0],
        ['Container Memory (%)', 72.3, 71.8],
        ['Pod Restart Count', 5, 0],
        ['Replica Count', 3, 3],
        ['Ingress Traffic (GB)', 145.2, 142.8],
        ['Egress Traffic (GB)', 89.7, 88.5],
        ['SSL Cert Days Left', 45, 44],
        ['DNS Lookup (ms)', 12.3, 11.8],
        ['TLS Handshake (ms)', 28.5, 27.0],
        ['CDN Cache Hit (%)', 88.2, 87.9],
        ['Origin Pull Rate (%)', 11.8, 12.1],
        ['Bandwidth Cost ($)', 2450, 2380],
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        ws.cell(row=r, column=2, value=row_data[1])
        ws.cell(row=r, column=3, value=row_data[2])

    # Set column widths for readability
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18

    # NO conditional formatting in initial state
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
