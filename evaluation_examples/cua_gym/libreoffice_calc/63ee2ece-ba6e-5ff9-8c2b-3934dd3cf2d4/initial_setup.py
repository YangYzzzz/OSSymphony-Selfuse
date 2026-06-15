"""
Initial Setup: Event Planning Dashboard - Raw Data
Task ID: calc_gpm_031
Domain: libreoffice_calc

Creates a spreadsheet with raw unformatted data for timeline, budget, and vendor
sections. The agent's task is to build the full formatted dashboard with charts.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_031'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'EventDash'

    # --- Raw data for Timeline section ---
    # Row 1: just a plain title, no merge, no formatting
    ws['A1'] = 'Corporate Retreat 2026 - Master Dashboard'

    # Row 3: section label
    ws['A3'] = 'Timeline'

    # Row 4: Headers (plain, unformatted)
    timeline_headers = ['Task', 'Owner', 'Start', 'End', 'Status', 'Days Left', 'Priority']
    for col, h in enumerate(timeline_headers, 1):
        ws.cell(row=4, column=col, value=h)

    # Rows 5-11: 7 timeline tasks with raw data
    timeline_data = [
        ['Venue Booking', 'Sarah Chen', '2026-05-01', '2026-05-10', 'Done', 39, 'High'],
        ['Catering Arrangements', 'Marcus Johnson', '2026-05-05', '2026-05-20', 'In Progress', 49, 'High'],
        ['Transport Logistics', 'Priya Patel', '2026-05-10', '2026-05-25', 'In Progress', 54, 'Medium'],
        ['Team Activities', 'David Kim', '2026-05-15', '2026-06-01', 'Not Started', 61, 'Medium'],
        ['Materials & Supplies', 'Emily Rodriguez', '2026-05-20', '2026-06-05', 'Not Started', 65, 'Low'],
        ['AV Setup', 'James Wilson', '2026-05-25', '2026-06-08', 'At Risk', 68, 'High'],
        ['Rehearsal', 'Lisa Thompson', '2026-06-01', '2026-06-10', 'Not Started', 70, 'Medium'],
    ]
    for r, row_data in enumerate(timeline_data, 5):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # --- Raw data for Budget section ---
    ws['A14'] = 'Budget Summary'

    budget_headers = ['Category', 'Budgeted', 'Spent', 'Remaining', '% Used']
    for col, h in enumerate(budget_headers, 1):
        ws.cell(row=15, column=col, value=h)

    budget_data = [
        ['Venue & Accommodation', 15000, 14200, 800, 0.947],
        ['Catering & Dining', 10000, 8500, 1500, 0.85],
        ['Transportation', 7000, 7350, -350, 1.05],
        ['Activities & Entertainment', 8000, 5200, 2800, 0.65],
        ['Materials & Equipment', 5000, 4800, 200, 0.96],
        ['AV & Technology', 5000, 3900, 1100, 0.78],
    ]
    for r, row_data in enumerate(budget_data, 16):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)

    # Total row (raw values, no formulas, no special border)
    ws.cell(row=22, column=1, value='Total')
    ws.cell(row=22, column=2, value=50000)
    ws.cell(row=22, column=3, value=43950)
    ws.cell(row=22, column=4, value=6050)
    ws.cell(row=22, column=5, value=0.879)

    # --- Raw data for Vendors section ---
    ws['A24'] = 'Vendors'

    vendor_headers = ['Vendor', 'Service', 'Contract', 'Paid', 'Balance', 'Status']
    for col, h in enumerate(vendor_headers, 1):
        ws.cell(row=25, column=col, value=h)

    vendor_data = [
        ['Grand Lakeside Resort', 'Venue & Rooms', 14200, 10000, 4200, 'Confirmed'],
        ['Epicurean Catering Co.', 'Food & Beverage', 8500, 5000, 3500, 'Confirmed'],
        ['Metro Transit Services', 'Bus & Shuttle', 7350, 7350, 0, 'Confirmed'],
        ['Adventure Works Inc.', 'Team Building', 5200, 2600, 2600, 'Pending'],
        ['TechVision AV', 'Sound & Projection', 3900, 1950, 1950, 'Pending'],
    ]
    for r, row_data in enumerate(vendor_data, 26):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
