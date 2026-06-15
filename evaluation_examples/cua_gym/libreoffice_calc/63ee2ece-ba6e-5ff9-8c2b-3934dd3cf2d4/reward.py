"""
Reward Script: Build a full event planning dashboard with timeline, budget breakdown,
vendor tracking, and charts.
Task ID: calc_gpm_031
Domain: libreoffice_calc
Scoring:
  Component 1 — Merged cells & title styling (A1:N1, A3:G3, A14:E14)    0.20 pts
  Component 2 — Days Left formulas (=Dx-TODAY())                          0.15 pts
  Component 3 — Data validations (Status + Vendor Status dropdowns)       0.15 pts
  Component 4 — Conditional formatting (4 rules)                          0.15 pts
  Component 5 — Charts (stacked bar + pie)                                0.20 pts
  Component 6 — Total row double borders (row 22)                         0.15 pts
  Total: 1.0
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_031'


def persist_app_state(domain: str):
    """Try to save any unsaved state in the GUI app."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists
    if 'EventDash' not in wb.sheetnames:
        print("CRITICAL: 'EventDash' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['EventDash']

    # =========================================================================
    # Component 1: Merged cells & title styling (0.20 points)
    # Initial has NO merges and plain A1. Golden has 3 merges + styled A1.
    # =========================================================================
    try:
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        comp1_score = 0.0

        # Check A1:N1 merge
        has_a1_merge = any('A1' in r and 'N1' in r for r in merged_ranges)
        if has_a1_merge:
            comp1_score += 0.04
            print("PASS: A1:N1 merged")
        else:
            print(f"FAIL: A1:N1 not merged. Merges found: {merged_ranges}")

        # Check A3:G3 merge
        has_a3_merge = any('A3' in r and 'G3' in r for r in merged_ranges)
        if has_a3_merge:
            comp1_score += 0.03
            print("PASS: A3:G3 merged")
        else:
            print(f"FAIL: A3:G3 not merged")

        # Check A14:E14 merge
        has_a14_merge = any('A14' in r and 'E14' in r for r in merged_ranges)
        if has_a14_merge:
            comp1_score += 0.03
            print("PASS: A14:E14 merged")
        else:
            print(f"FAIL: A14:E14 not merged")

        # Check A1 font: 18pt bold white
        cell_a1 = ws['A1']
        a1_bold = cell_a1.font.bold is True
        a1_size = cell_a1.font.size is not None and cell_a1.font.size >= 16
        a1_white_font = False
        try:
            color_rgb = cell_a1.font.color.rgb if cell_a1.font.color else None
            if color_rgb and 'FFFFFF' in str(color_rgb).upper():
                a1_white_font = True
        except Exception:
            pass

        if a1_bold and a1_size:
            comp1_score += 0.05
            print(f"PASS: A1 font bold={a1_bold}, size={cell_a1.font.size}")
        else:
            print(f"FAIL: A1 font bold={a1_bold}, size={cell_a1.font.size} (expected bold + >=16pt)")

        # Check A1 dark olive fill (RGB 85,107,47 => hex 556B2F)
        a1_fill_ok = False
        try:
            fg_rgb = cell_a1.fill.fgColor.rgb if cell_a1.fill.fgColor else None
            if fg_rgb and '556B2F' in str(fg_rgb).upper():
                a1_fill_ok = True
        except Exception:
            pass

        if a1_fill_ok:
            comp1_score += 0.05
            print(f"PASS: A1 fill is dark olive (556B2F)")
        else:
            fg_val = None
            try:
                fg_val = cell_a1.fill.fgColor.rgb
            except Exception:
                pass
            print(f"FAIL: A1 fill expected dark olive (556B2F), found: {fg_val}")

        total_score += comp1_score
        print(f"Component 1 subtotal: {comp1_score}/0.20")

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Days Left formulas =Dx-TODAY() in F5:F11 (0.15 points)
    # Initial has static integers. Golden has formulas.
    # =========================================================================
    try:
        formula_count = 0
        for row in range(5, 12):  # F5 through F11
            val = ws.cell(row=row, column=6).value  # column F
            if isinstance(val, str) and 'TODAY()' in val.upper():
                formula_count += 1

        if formula_count == 7:
            print(f"PASS: Component 2 — All 7 Days Left cells have TODAY() formulas ({formula_count}/7)")
            total_score += 0.15
        elif formula_count >= 4:
            partial = round(0.15 * formula_count / 7, 3)
            print(f"PARTIAL: Component 2 — {formula_count}/7 Days Left formulas found (+{partial})")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {formula_count}/7 Days Left cells have TODAY() formulas")

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Data validations (0.15 points)
    # Initial has 0 validations. Golden has 2.
    # =========================================================================
    try:
        dvs = ws.data_validations.dataValidation if ws.data_validations else []
        dv_count = len(dvs)
        comp3_score = 0.0

        # Check for Status validation (E5:E11 with Done/In Progress/Not Started/At Risk)
        status_dv_found = False
        vendor_dv_found = False
        for dv in dvs:
            formula_str = str(dv.formula1) if dv.formula1 else ''
            sqref_str = str(dv.sqref) if dv.sqref else ''

            if dv.type == 'list' and 'Done' in formula_str and 'E' in sqref_str:
                status_dv_found = True
            if dv.type == 'list' and 'Confirmed' in formula_str and 'F' in sqref_str:
                vendor_dv_found = True

        if status_dv_found:
            comp3_score += 0.08
            print("PASS: Status dropdown validation found on E column")
        else:
            print(f"FAIL: Status dropdown validation not found (dvs={dv_count})")

        if vendor_dv_found:
            comp3_score += 0.07
            print("PASS: Vendor Status dropdown validation found on F column")
        else:
            print(f"FAIL: Vendor Status dropdown validation not found")

        total_score += comp3_score
        print(f"Component 3 subtotal: {comp3_score}/0.15")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: Conditional formatting (0.15 points)
    # Initial has 0 CF rules. Golden has 4 CF ranges with multiple rules.
    # =========================================================================
    try:
        cf_list = list(ws.conditional_formatting)
        cf_count = len(cf_list)
        comp4_score = 0.0

        # Check for status conditional formatting on E5:E11
        has_status_cf = False
        has_days_cf = False
        has_databars = False
        has_remaining_cf = False

        for cf in cf_list:
            cf_range = str(cf)
            for rule in cf.rules:
                rule_type = rule.type
                # Status CF: cellIs rules on E5:E11
                if 'E5' in cf_range and rule_type == 'cellIs':
                    has_status_cf = True
                # Days Left CF: cellIs lessThan 0 on F5:F11
                if 'F5' in cf_range and rule_type == 'cellIs':
                    has_days_cf = True
                # Data bars on % Used (E16:E21)
                if 'E16' in cf_range and rule_type == 'dataBar':
                    has_databars = True
                # Remaining negative red (D16:D21)
                if 'D16' in cf_range and rule_type == 'cellIs':
                    has_remaining_cf = True

        if has_status_cf:
            comp4_score += 0.05
            print("PASS: Status conditional formatting found (E5:E11)")
        else:
            print("FAIL: Status conditional formatting not found")

        if has_days_cf:
            comp4_score += 0.03
            print("PASS: Days Left negative CF found (F5:F11)")
        else:
            print("FAIL: Days Left negative CF not found")

        if has_databars:
            comp4_score += 0.04
            print("PASS: Data bars found on % Used (E16:E21)")
        else:
            print("FAIL: Data bars not found on % Used column")

        if has_remaining_cf:
            comp4_score += 0.03
            print("PASS: Remaining negative red CF found (D16:D21)")
        else:
            print("FAIL: Remaining negative red CF not found")

        total_score += comp4_score
        print(f"Component 4 subtotal: {comp4_score}/0.15")

    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # =========================================================================
    # Component 5: Charts — stacked bar + pie (0.20 points)
    # Initial has 0 charts. Golden has 2.
    # =========================================================================
    try:
        charts = ws._charts
        comp5_score = 0.0

        # Check for stacked bar chart titled "Budget Tracking"
        bar_chart_found = False
        pie_chart_found = False

        for ch in charts:
            ch_class = type(ch).__name__

            # Extract title text
            title_text = ''
            try:
                if ch.title and ch.title.tx and ch.title.tx.rich:
                    for p in ch.title.tx.rich.p:
                        for r in p.r:
                            title_text += r.t
            except Exception:
                pass

            if ch_class == 'BarChart':
                # Check if stacked
                grouping = getattr(ch, 'grouping', None)
                if grouping == 'stacked' and 'Budget' in title_text:
                    bar_chart_found = True
                    print(f"PASS: Stacked bar chart found, title='{title_text}', grouping={grouping}")
                elif ch_class == 'BarChart':
                    # Partial: bar chart exists but may not be stacked or titled correctly
                    bar_chart_found = True
                    print(f"PASS: Bar chart found (title='{title_text}', grouping={grouping})")

            if ch_class == 'PieChart':
                pie_chart_found = True
                print(f"PASS: Pie chart found, title='{title_text}'")

        if bar_chart_found:
            comp5_score += 0.10
        else:
            print("FAIL: No bar chart found")

        if pie_chart_found:
            comp5_score += 0.10
        else:
            print("FAIL: No pie chart found")

        total_score += comp5_score
        print(f"Component 5 subtotal: {comp5_score}/0.20")

    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # =========================================================================
    # Component 6: Total row double borders on row 22 (0.15 points)
    # Initial has no special borders. Golden has double borders top+bottom.
    # =========================================================================
    try:
        comp6_score = 0.0
        double_border_count = 0

        for col_letter in ['A', 'B', 'C', 'D', 'E']:
            cell = ws[f'{col_letter}22']
            top_style = cell.border.top.style if cell.border.top else None
            bottom_style = cell.border.bottom.style if cell.border.bottom else None
            if top_style == 'double' or bottom_style == 'double':
                double_border_count += 1

        if double_border_count >= 4:
            comp6_score = 0.15
            print(f"PASS: Component 6 — {double_border_count}/5 budget total cells have double borders")
        elif double_border_count >= 2:
            comp6_score = round(0.15 * double_border_count / 5, 3)
            print(f"PARTIAL: Component 6 — {double_border_count}/5 cells with double borders (+{comp6_score})")
        else:
            print(f"FAIL: Component 6 — Only {double_border_count}/5 total row cells have double borders")

        total_score += comp6_score
        print(f"Component 6 subtotal: {comp6_score}/0.15")

    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # =========================================================================
    # Final score
    # =========================================================================
    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
