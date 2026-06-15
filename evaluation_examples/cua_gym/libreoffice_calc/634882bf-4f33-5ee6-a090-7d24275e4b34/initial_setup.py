"""
Initial Setup: Sales Commission Calculator
Task ID: calc_wf_015
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_015'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Commissions ---
    ws = wb.active
    ws.title = 'Commissions'

    # Headers
    headers = ['Rep Name', 'Total Sales', 'Commission', 'Rank']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font_white = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 15 Sales Representatives with realistic names and sales amounts ($5,000 - $50,000)
    reps = [
        ('Sarah Chen', 47500),
        ('Marcus Johnson', 12800),
        ('Elena Rodriguez', 31200),
        ('David Kim', 8750),
        ('Priya Patel', 22400),
        ('James O\'Brien', 15600),
        ('Aisha Mohammed', 41300),
        ('Tyler Washington', 6200),
        ('Mei-Lin Zhang', 28900),
        ('Robert Hernandez', 50000),
        ('Olivia Foster', 9400),
        ('Carlos Gutierrez', 18750),
        ('Hannah Johansson', 35600),
        ('Nathan Brooks', 5100),
        ('Fatima Al-Rashid', 24300),
    ]

    for r, (name, sales) in enumerate(reps, 2):
        # Rep Name
        name_cell = ws.cell(row=r, column=1, value=name)
        name_cell.font = Font(name='Calibri', size=11)
        name_cell.border = thin_border

        # Total Sales
        sales_cell = ws.cell(row=r, column=2, value=sales)
        sales_cell.number_format = '$#,##0.00'
        sales_cell.font = Font(name='Calibri', size=11)
        sales_cell.border = thin_border

        # Commission column - LEFT EMPTY (task requires creating formulas)
        comm_cell = ws.cell(row=r, column=3)
        comm_cell.border = thin_border
        comm_cell.number_format = '$#,##0.00'

        # Rank column - LEFT EMPTY (task requires adding RANK function)
        rank_cell = ws.cell(row=r, column=4)
        rank_cell.border = thin_border
        rank_cell.alignment = Alignment(horizontal='center')

    # Set column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 10

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
