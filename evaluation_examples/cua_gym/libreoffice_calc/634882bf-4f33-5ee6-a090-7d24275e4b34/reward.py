"""
Reward Script: Sales Commission Calculator
Task ID: calc_wf_015
Domain: libreoffice_calc
Scoring:
  Component 1: Commission formulas in C2:C16 (tiered IF)       — 0.35 points
  Component 2: RANK formulas in D2:D16                         — 0.20 points
  Component 3: Data bar conditional formatting on C2:C16       — 0.20 points
  Component 4: Pie chart present with commission distribution   — 0.25 points
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_015'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Commissions' sheet must exist
    if 'Commissions' not in wb.sheetnames:
        print("FAIL: 'Commissions' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Commissions']

    # Component 1: Commission formulas in C2:C16 (0.35 points)
    # The task requires tiered commission: 5% up to $10K, 8% for $10K-$25K, 12% above $25K
    # Golden uses nested IF formula. We check that C2:C16 contain formulas with IF and tiered logic.
    try:
        commission_formula_count = 0
        for r in range(2, 17):
            val = ws.cell(row=r, column=3).value
            if val is not None and isinstance(val, str) and val.startswith('='):
                # Check it references the tiered structure (IF with 0.05, 0.08, 0.12 rates)
                formula_upper = val.upper().replace(' ', '')
                has_if = 'IF(' in formula_upper
                has_tiers = ('0.05' in val and '0.08' in val and '0.12' in val)
                if has_if and has_tiers:
                    commission_formula_count += 1
        if commission_formula_count == 15:
            print(f"PASS: Component 1 — All 15 commission formulas found with tiered IF logic (0.35 pts)")
            total_score += 0.35
        elif commission_formula_count >= 10:
            partial = round(0.35 * commission_formula_count / 15, 2)
            print(f"PARTIAL: Component 1 — {commission_formula_count}/15 commission formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {commission_formula_count}/15 valid tiered commission formulas")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: RANK formulas in D2:D16 (0.20 points)
    # The task requires ranking reps by total commission.
    try:
        rank_formula_count = 0
        for r in range(2, 17):
            val = ws.cell(row=r, column=4).value
            if val is not None and isinstance(val, str) and val.startswith('='):
                formula_upper = val.upper().replace(' ', '')
                if 'RANK(' in formula_upper or 'RANK.EQ(' in formula_upper or 'RANK.AVG(' in formula_upper:
                    rank_formula_count += 1
        if rank_formula_count == 15:
            print(f"PASS: Component 2 — All 15 RANK formulas found in column D (0.20 pts)")
            total_score += 0.20
        elif rank_formula_count >= 10:
            partial = round(0.20 * rank_formula_count / 15, 2)
            print(f"PARTIAL: Component 2 — {rank_formula_count}/15 RANK formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {rank_formula_count}/15 RANK formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Data bar conditional formatting on C2:C16 (0.20 points)
    # The task requires data bars on the commission column.
    try:
        has_data_bar = False
        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            for rule in cf.rules:
                if rule.type == 'dataBar':
                    # Check that the range covers at least some of C2:C16
                    if 'C' in cf_range:
                        has_data_bar = True
                        break
            if has_data_bar:
                break
        if has_data_bar:
            print(f"PASS: Component 3 — Data bar conditional formatting found on commission column (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 3 — No data bar conditional formatting found on column C")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Pie chart showing commission distribution (0.25 points)
    # The task requires a pie chart of commission distribution.
    try:
        pie_chart_found = False
        # Check all sheets for pie charts
        for sname in wb.sheetnames:
            sheet = wb[sname]
            for chart in sheet._charts:
                if isinstance(chart, openpyxl.chart.PieChart):
                    pie_chart_found = True
                    break
            if pie_chart_found:
                break

        if pie_chart_found:
            print(f"PASS: Component 4 — Pie chart found (0.25 pts)")
            total_score += 0.25
        else:
            # Also check for any chart type as partial credit
            any_chart = False
            for sname in wb.sheetnames:
                sheet = wb[sname]
                if len(sheet._charts) > 0:
                    any_chart = True
                    break
            if any_chart:
                print(f"PARTIAL: Component 4 — Chart found but not a pie chart (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 4 — No charts found in any sheet")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
