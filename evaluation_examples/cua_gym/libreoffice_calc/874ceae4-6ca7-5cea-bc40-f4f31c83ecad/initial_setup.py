"""
Initial Setup: Chart data range change - Monthly revenue line chart using A1:B7 (Jan-June only)
Task ID: calc_chart_data_range_change_031
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.chart import LineChart, Reference

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_data_range_change_031'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Monthly ---
    ws = wb.active
    ws.title = 'Monthly'

    # Headers
    ws['A1'] = 'Month'
    ws['B1'] = 'Revenue'

    # Data rows 2-8 (January through July)
    monthly_data = [
        ('January',  85000),
        ('February', 91000),
        ('March',    102000),
        ('April',    98000),
        ('May',      115000),
        ('June',     128000),
        ('July',     142000),  # newly added row — NOT yet in the chart range
    ]

    for r, (month, revenue) in enumerate(monthly_data, 2):
        ws.cell(row=r, column=1, value=month)
        ws.cell(row=r, column=2, value=revenue)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14

    # --- Line Chart covering only A1:B7 (header + Jan through June, NOT July) ---
    chart = LineChart()
    chart.title = 'Monthly Revenue'
    chart.style = 10
    chart.y_axis.title = 'Revenue ($)'
    chart.x_axis.title = 'Month'

    # Data series: B1:B7 (Revenue header + Jan-Jun values = rows 1-7)
    data_ref = Reference(ws, min_col=2, min_row=1, max_row=7)
    chart.add_data(data_ref, titles_from_data=True)

    # Categories: A2:A7 (month names Jan-Jun only)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=7)
    chart.set_categories(cats_ref)

    chart.shape = 4
    ws.add_chart(chart, 'D2')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Chart uses data range A1:B7 (January through June only)')
    print('Row 8 (July=142000) is present but NOT included in chart range')


create_initial()
