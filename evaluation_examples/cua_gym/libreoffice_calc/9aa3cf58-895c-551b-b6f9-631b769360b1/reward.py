"""
Reward Script: Build a resource utilization tracker
Task ID: calc_ops_project_tracking_resource_014
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: SUMIF formulas in B2:B7 to sum allocated hours per team member    — 0.30 pts
  Component 2: Capacity values in C2:C7 set to 160                                — 0.20 pts
  Component 3: Utilization ratio formulas in D2:D7 (=Bx/Cx) with % number format  — 0.20 pts
  Component 4: IF status formulas in E2:E7 (OVER-ALLOCATED / At Capacity / Available) — 0.20 pts
  Component 5: Conditional formatting on D2:D7 (red >100%, green <=100%)          — 0.10 pts
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_project_tracking_resource_014'

TEAM_MEMBERS = ['Alex', 'Maria', 'James', 'Sarah', 'Tom', 'Linda']


def verify_task(file_path):
    """
    Verify resource utilization tracker task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Gate: load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Gate: check required sheets exist
    if 'Utilization' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Utilization' not found")
        print("REWARD: 0.0")
        return 0.0

    ws_util = wb['Utilization']

    # -------------------------------------------------------------------------
    # Component 1: SUMIF formulas in B2:B7 to sum allocated hours per team member
    # (0.30 points)
    # Each cell must contain a SUMIF formula referencing TaskAllocation!C and D columns
    # -------------------------------------------------------------------------
    try:
        sumif_pass = 0
        for row in range(2, 8):
            cell_val = ws_util.cell(row=row, column=2).value  # Column B
            if cell_val and isinstance(cell_val, str):
                val_upper = cell_val.upper().replace(' ', '')
                # Must contain SUMIF referencing TaskAllocation sheet
                if 'SUMIF' in val_upper and 'TASKALLOCATION' in val_upper:
                    sumif_pass += 1
                    print(f"PASS: B{row} has SUMIF formula: {cell_val}")
                else:
                    print(f"FAIL: B{row} formula does not reference SUMIF on TaskAllocation: {cell_val}")
            else:
                print(f"FAIL: B{row} is empty or not a formula: {repr(cell_val)}")

        if sumif_pass == 6:
            print(f"PASS: Component 1 — All 6 SUMIF formulas present in B2:B7 (0.30 pts)")
            total_score += 0.30
        elif sumif_pass >= 3:
            partial = round(0.30 * sumif_pass / 6, 4)
            print(f"PARTIAL: Component 1 — {sumif_pass}/6 SUMIF formulas present ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {sumif_pass}/6 SUMIF formulas present (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Capacity values in C2:C7 set to 160
    # (0.20 points)
    # -------------------------------------------------------------------------
    try:
        capacity_pass = 0
        for row in range(2, 8):
            cell_val = ws_util.cell(row=row, column=3).value  # Column C
            # Accept integer 160, float 160.0, or string '160'
            try:
                numeric_val = float(cell_val) if cell_val is not None else None
            except (TypeError, ValueError):
                numeric_val = None

            if numeric_val is not None and abs(numeric_val - 160.0) < 0.01:
                capacity_pass += 1
                print(f"PASS: C{row} capacity = 160")
            else:
                print(f"FAIL: C{row} expected 160, found: {repr(cell_val)}")

        if capacity_pass == 6:
            print(f"PASS: Component 2 — All 6 capacity values are 160 in C2:C7 (0.20 pts)")
            total_score += 0.20
        elif capacity_pass >= 3:
            partial = round(0.20 * capacity_pass / 6, 4)
            print(f"PARTIAL: Component 2 — {capacity_pass}/6 capacity values are 160 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {capacity_pass}/6 capacity values are 160 (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Utilization ratio formulas in D2:D7 (=Bx/Cx) with % number format
    # (0.20 points)
    # Formula must be division (B/C) and cell must be formatted as percentage
    # -------------------------------------------------------------------------
    try:
        util_formula_pass = 0
        util_format_pass = 0
        for row in range(2, 8):
            cell = ws_util.cell(row=row, column=4)  # Column D
            cell_val = cell.value
            # Check for division formula (Bx/Cx pattern)
            if cell_val and isinstance(cell_val, str):
                val_upper = cell_val.upper().replace(' ', '')
                # Pattern: =B2/C2 or similar ratio formula
                if '/' in val_upper and val_upper.startswith('=B') and '/C' in val_upper:
                    util_formula_pass += 1
                    print(f"PASS: D{row} has utilization formula: {cell_val}")
                else:
                    print(f"FAIL: D{row} formula is not a B/C division: {cell_val}")
            else:
                print(f"FAIL: D{row} is empty or not a formula: {repr(cell_val)}")

            # Check percentage number format
            num_fmt = cell.number_format
            if num_fmt and '%' in str(num_fmt):
                util_format_pass += 1
                print(f"PASS: D{row} has percentage number format: {num_fmt}")
            else:
                print(f"FAIL: D{row} number format not percentage: {repr(num_fmt)}")

        # Score: formula must be present AND formatted as %
        formula_ok = util_formula_pass == 6
        format_ok = util_format_pass == 6
        if formula_ok and format_ok:
            print(f"PASS: Component 3 — All 6 utilization formulas with % format (0.20 pts)")
            total_score += 0.20
        elif formula_ok or format_ok:
            print(f"PARTIAL: Component 3 — formulas={util_formula_pass}/6, formats={util_format_pass}/6 (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — formulas={util_formula_pass}/6, formats={util_format_pass}/6 (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: IF status formulas in E2:E7 — OVER-ALLOCATED / At Capacity / Available
    # (0.20 points)
    # Must use IF formula with three status values based on utilization percentage
    # -------------------------------------------------------------------------
    try:
        status_pass = 0
        for row in range(2, 8):
            cell_val = ws_util.cell(row=row, column=5).value  # Column E
            if cell_val and isinstance(cell_val, str):
                val_upper = cell_val.upper().replace(' ', '')
                # Must be an IF formula containing OVER-ALLOCATED
                if val_upper.startswith('=IF') and 'OVER-ALLOCATED' in val_upper:
                    # Also check it includes "At Capacity" and "Available"
                    val_orig = cell_val.upper()
                    if 'AT CAPACITY' in val_orig and 'AVAILABLE' in val_orig:
                        status_pass += 1
                        print(f"PASS: E{row} has complete IF status formula: {cell_val}")
                    else:
                        print(f"FAIL: E{row} IF formula missing some status values: {cell_val}")
                else:
                    print(f"FAIL: E{row} formula is not IF with OVER-ALLOCATED: {cell_val}")
            else:
                print(f"FAIL: E{row} is empty or not a formula: {repr(cell_val)}")

        if status_pass == 6:
            print(f"PASS: Component 4 — All 6 IF status formulas present in E2:E7 (0.20 pts)")
            total_score += 0.20
        elif status_pass >= 3:
            partial = round(0.20 * status_pass / 6, 4)
            print(f"PARTIAL: Component 4 — {status_pass}/6 IF status formulas present ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {status_pass}/6 IF status formulas present (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -------------------------------------------------------------------------
    # Component 5: Conditional formatting on D2:D7
    # (0.10 points)
    # Must have at least 2 CF rules on D2:D7: red fill for >100%, green for <=100%
    # -------------------------------------------------------------------------
    try:
        cf_found = False
        has_red_rule = False
        has_green_rule = False

        for cf_range_obj in ws_util.conditional_formatting:
            cf_range_str = str(cf_range_obj)
            rules_list = ws_util.conditional_formatting._cf_rules[cf_range_obj]
            # Check if it covers D2:D7 (or a range that includes it)
            if 'D' in cf_range_str.upper() and '2' in cf_range_str:
                cf_found = True
                for rule in rules_list:
                    rule_type = getattr(rule, 'type', None)
                    dxf = getattr(rule, 'dxf', None)
                    if dxf and dxf.fill:
                        try:
                            fill_color = dxf.fill.fgColor.rgb
                            # Red fill: ARGB containing FF0000
                            if 'FF0000' in fill_color.upper():
                                has_red_rule = True
                                print(f"PASS: CF has red fill rule (color: {fill_color})")
                            # Green fill: ARGB containing 00FF00
                            elif '00FF00' in fill_color.upper():
                                has_green_rule = True
                                print(f"PASS: CF has green fill rule (color: {fill_color})")
                        except Exception:
                            pass

        if cf_found and has_red_rule and has_green_rule:
            print(f"PASS: Component 5 — Conditional formatting on D2:D7 with red and green (0.10 pts)")
            total_score += 0.10
        elif cf_found and (has_red_rule or has_green_rule):
            print(f"PARTIAL: Component 5 — CF found but only one color rule (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — Conditional formatting not found or incomplete on D2:D7")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # -------------------------------------------------------------------------
    # Final score
    # -------------------------------------------------------------------------
    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
