"""
Initial Setup: Resource Utilization Tracker - TaskAllocation and Utilization sheets
Task ID: calc_ops_project_tracking_resource_014
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_project_tracking_resource_014'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: TaskAllocation ---
    ws1 = wb.active
    ws1.title = 'TaskAllocation'

    # Headers
    headers = ['Task ID', 'Task Name', 'Assigned To', 'Estimated Hours', 'Month']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFFFF')

    # 60 task allocation rows for 6 team members
    team_members = ['Alex', 'Maria', 'James', 'Sarah', 'Tom', 'Linda']
    current_month = '2025-03'

    task_data = [
        # Alex - tasks (10 tasks)
        ('T-001', 'Backend API Development', 'Alex', 18, current_month),
        ('T-002', 'Database Schema Optimization', 'Alex', 12, current_month),
        ('T-003', 'Authentication Module', 'Alex', 20, current_month),
        ('T-004', 'Unit Test Coverage', 'Alex', 8, current_month),
        ('T-005', 'API Documentation', 'Alex', 6, current_month),
        ('T-006', 'Performance Profiling', 'Alex', 14, current_month),
        ('T-007', 'CI/CD Pipeline Setup', 'Alex', 10, current_month),
        ('T-008', 'Code Review - Sprint 12', 'Alex', 8, current_month),
        ('T-009', 'Security Audit Fix', 'Alex', 16, current_month),
        ('T-010', 'Microservice Refactor', 'Alex', 22, current_month),

        # Maria - tasks (10 tasks)
        ('T-011', 'UI Component Library', 'Maria', 20, current_month),
        ('T-012', 'Dashboard Redesign', 'Maria', 15, current_month),
        ('T-013', 'Mobile Responsiveness', 'Maria', 12, current_month),
        ('T-014', 'Accessibility Improvements', 'Maria', 8, current_month),
        ('T-015', 'CSS Framework Migration', 'Maria', 18, current_month),
        ('T-016', 'User Testing Sessions', 'Maria', 6, current_month),
        ('T-017', 'Design System Documentation', 'Maria', 10, current_month),
        ('T-018', 'Animation Prototyping', 'Maria', 7, current_month),
        ('T-019', 'Icon Set Creation', 'Maria', 9, current_month),
        ('T-020', 'Frontend Performance Audit', 'Maria', 11, current_month),

        # James - tasks (10 tasks)
        ('T-021', 'Data Pipeline Architecture', 'James', 24, current_month),
        ('T-022', 'ETL Process Optimization', 'James', 16, current_month),
        ('T-023', 'ML Model Integration', 'James', 20, current_month),
        ('T-024', 'Data Quality Checks', 'James', 10, current_month),
        ('T-025', 'Analytics Dashboard', 'James', 14, current_month),
        ('T-026', 'Real-time Data Streaming', 'James', 18, current_month),
        ('T-027', 'Reporting Automation', 'James', 8, current_month),
        ('T-028', 'Data Warehouse Schema', 'James', 12, current_month),
        ('T-029', 'KPI Metrics Definition', 'James', 6, current_month),
        ('T-030', 'Stakeholder Reports', 'James', 10, current_month),

        # Sarah - tasks (10 tasks)
        ('T-031', 'Product Roadmap Planning', 'Sarah', 14, current_month),
        ('T-032', 'Sprint Retrospective', 'Sarah', 4, current_month),
        ('T-033', 'Feature Requirements Analysis', 'Sarah', 16, current_month),
        ('T-034', 'Customer Feedback Review', 'Sarah', 8, current_month),
        ('T-035', 'Competitive Analysis', 'Sarah', 12, current_month),
        ('T-036', 'OKR Alignment Meeting', 'Sarah', 6, current_month),
        ('T-037', 'Release Planning', 'Sarah', 10, current_month),
        ('T-038', 'User Story Writing', 'Sarah', 18, current_month),
        ('T-039', 'Stakeholder Demos', 'Sarah', 8, current_month),
        ('T-040', 'Market Research', 'Sarah', 14, current_month),

        # Tom - tasks (10 tasks)
        ('T-041', 'Cloud Infrastructure Setup', 'Tom', 22, current_month),
        ('T-042', 'Kubernetes Cluster Migration', 'Tom', 28, current_month),
        ('T-043', 'Monitoring & Alerting', 'Tom', 15, current_month),
        ('T-044', 'Disaster Recovery Plan', 'Tom', 12, current_month),
        ('T-045', 'Cost Optimization Review', 'Tom', 8, current_month),
        ('T-046', 'Security Patch Deployment', 'Tom', 10, current_month),
        ('T-047', 'Load Balancer Configuration', 'Tom', 14, current_month),
        ('T-048', 'Database Backup Automation', 'Tom', 9, current_month),
        ('T-049', 'SSL Certificate Renewal', 'Tom', 4, current_month),
        ('T-050', 'Network Firewall Rules', 'Tom', 16, current_month),

        # Linda - tasks (10 tasks)
        ('T-051', 'QA Test Plan Creation', 'Linda', 14, current_month),
        ('T-052', 'Regression Test Suite', 'Linda', 20, current_month),
        ('T-053', 'Bug Triage and Prioritization', 'Linda', 8, current_month),
        ('T-054', 'Test Automation Framework', 'Linda', 18, current_month),
        ('T-055', 'Performance Testing', 'Linda', 12, current_month),
        ('T-056', 'API Contract Testing', 'Linda', 10, current_month),
        ('T-057', 'Cross-browser Testing', 'Linda', 8, current_month),
        ('T-058', 'Defect Report Analysis', 'Linda', 6, current_month),
        ('T-059', 'QA Process Documentation', 'Linda', 7, current_month),
        ('T-060', 'User Acceptance Testing', 'Linda', 15, current_month),
    ]

    for r, row_data in enumerate(task_data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Column widths for readability
    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 35
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 18
    ws1.column_dimensions['E'].width = 12

    # --- Sheet 2: Utilization ---
    ws2 = wb.create_sheet('Utilization')

    # Headers
    util_headers = ['Team Member', 'Allocated Hours', 'Capacity Hours', 'Utilization %', 'Status']
    for col, h in enumerate(util_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFFFF')

    # Team member names in column A (rows 2-7), columns B-E are EMPTY (no formulas yet)
    for r, name in enumerate(team_members, 2):
        ws2.cell(row=r, column=1, value=name)
        # B, C, D, E columns intentionally left empty — user must fill these in

    # Column widths
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 15
    ws2.column_dimensions['E'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  - TaskAllocation sheet: 60 task rows, 6 team members')
    print(f'  - Utilization sheet: 6 team member rows, B-E columns empty')


create_initial()
