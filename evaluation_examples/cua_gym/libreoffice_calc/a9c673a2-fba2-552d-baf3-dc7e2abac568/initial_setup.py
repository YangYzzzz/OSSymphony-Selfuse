"""
Initial Setup: Set up data validation on hire date column
Task ID: calc_hr_030
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
from datetime import date, timedelta

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_030'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: NewHires ---
    ws = wb.active
    ws.title = 'NewHires'

    # Headers
    headers = ['Employee', 'Hire Date', 'Department']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)

    # Employee data - 9 rows (A2:A10, B2:B10, C2:C10)
    today = date.today()
    data = [
        ['Sarah Chen',       date(2025, 3, 15),  'Engineering'],
        ['Marcus Johnson',   date(2025, 6, 1),   'Marketing'],
        ['Priya Patel',      date(2024, 11, 20), 'Finance'],
        ['James O\'Brien',   date(2025, 1, 8),   'Human Resources'],
        ['Aisha Williams',   date(2025, 9, 12),  'Engineering'],
        ['Carlos Rivera',    date(2024, 7, 22),  'Sales'],
        ['Emily Nakamura',   date(2025, 4, 3),   'Product'],
        ['David Kim',        date(2025, 8, 18),  'Operations'],
        ['Fatima Al-Hassan', date(2024, 12, 5),  'Legal'],
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        ws.cell(row=r, column=2, value=row_data[1])
        ws.cell(row=r, column=2).number_format = 'yyyy-mm-dd'
        ws.cell(row=r, column=3, value=row_data[2])

    # Set column widths for readability
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18

    # NO data validation on B2:B10 - that is the task for the agent

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
