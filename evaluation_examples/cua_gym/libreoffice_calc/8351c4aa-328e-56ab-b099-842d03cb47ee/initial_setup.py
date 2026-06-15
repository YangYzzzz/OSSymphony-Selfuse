"""
Initial Setup: Create named ranges for quarterly revenue data and summary formulas
Task ID: calc_nrv_015
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_015'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Headers
    ws.cell(row=1, column=1, value="Month")
    ws.cell(row=1, column=2, value="Revenue")
    ws.cell(row=1, column=5, value="Quarter Summary")

    # Monthly revenue data (Jan-Dec) in rows 2-13
    months_data = [
        ("January",   45230),
        ("February",  38750),
        ("March",     52100),
        ("April",     41680),
        ("May",       47920),
        ("June",      53410),
        ("July",      39870),
        ("August",    44560),
        ("September", 51230),
        ("October",   48370),
        ("November",  55890),
        ("December",  62150),
    ]

    for r, (month, revenue) in enumerate(months_data, 2):
        ws.cell(row=r, column=1, value=month)
        ws.cell(row=r, column=2, value=revenue)

    # E2:E5 intentionally left empty - task requires creating formulas here
    # No named ranges defined - task requires creating them

    # Adjust column widths for readability
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["E"].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
