"""
Reward Script: Create named ranges for quarterly data and summary formulas
Task ID: calc_nrv_015
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Named ranges Q1-Q4 exist with correct cell references
  Component 2 (0.3): E2:E5 contain SUM formulas referencing Q1-Q4
  Component 3 (0.3): Formulas produce correct computed values
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_015'

# Expected named ranges: name -> expected reference (without sheet prefix)
EXPECTED_RANGES = {
    'Q1': '$B$2:$B$4',
    'Q2': '$B$5:$B$7',
    'Q3': '$B$8:$B$10',
    'Q4': '$B$11:$B$13',
}

# Expected formulas in E2:E5
EXPECTED_FORMULAS = {
    'E2': 'Q1',
    'E3': 'Q2',
    'E4': 'Q3',
    'E5': 'Q4',
}

# Expected computed values (sum of the respective quarterly revenues)
# Q1: 45230 + 38750 + 52100 = 136080
# Q2: 41680 + 47920 + 53410 = 143010
# Q3: 39870 + 44560 + 51230 = 135660
# Q4: 48370 + 55890 + 62150 = 166410
EXPECTED_VALUES = {
    'E2': 136080,
    'E3': 143010,
    'E4': 135660,
    'E5': 166410,
}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Component 1: Named ranges Q1-Q4 exist with correct references (0.4 points)
    # 0.1 per correct named range
    try:
        named_range_score = 0.0
        for name, expected_ref in EXPECTED_RANGES.items():
            if name in wb.defined_names:
                dn = wb.defined_names[name]
                actual_ref = dn.attr_text
                # The reference may include sheet name like "Sheet1!$B$2:$B$4"
                # Strip sheet prefix for comparison
                ref_part = actual_ref.split('!')[-1] if '!' in actual_ref else actual_ref
                if ref_part == expected_ref:
                    print(f"PASS: Named range '{name}' = {actual_ref} (correct)")
                    named_range_score += 0.1
                else:
                    print(f"FAIL: Named range '{name}' = {actual_ref}, expected ...!{expected_ref}")
            else:
                print(f"FAIL: Named range '{name}' not found")

        if named_range_score > 0:
            total_score += named_range_score
            print(f"  Component 1 subtotal: {named_range_score}/0.4")
    except Exception as e:
        print(f"ERROR: Component 1 (named ranges) - {e}")

    # Component 2: E2:E5 contain SUM formulas referencing Q1-Q4 (0.3 points)
    # 0.075 per correct formula
    try:
        formula_score = 0.0
        for cell_ref, range_name in EXPECTED_FORMULAS.items():
            val = ws[cell_ref].value
            if val is not None and isinstance(val, str):
                # Normalize: uppercase, remove spaces
                norm_val = val.upper().replace(' ', '')
                expected_pattern = f'=SUM({range_name})'.upper()
                if norm_val == expected_pattern:
                    print(f"PASS: {cell_ref} formula = {val} (correct)")
                    formula_score += 0.075
                else:
                    print(f"FAIL: {cell_ref} formula = {val}, expected =SUM({range_name})")
            else:
                print(f"FAIL: {cell_ref} value = {val}, expected =SUM({range_name})")

        if formula_score > 0:
            total_score += formula_score
            print(f"  Component 2 subtotal: {formula_score}/0.3")
    except Exception as e:
        print(f"ERROR: Component 2 (formulas) - {e}")

    # Component 3: Named ranges reference correct data that sums to expected values (0.3 points)
    # Verify by reading the actual cell values in the named range and computing the sum
    # 0.075 per quarter with correct sum
    try:
        import re
        value_score = 0.0

        # Map cell_ref -> (range_name, expected_sum)
        quarter_checks = {
            'E2': ('Q1', EXPECTED_VALUES['E2']),
            'E3': ('Q2', EXPECTED_VALUES['E3']),
            'E4': ('Q3', EXPECTED_VALUES['E4']),
            'E5': ('Q4', EXPECTED_VALUES['E5']),
        }

        for cell_ref, (range_name, expected_sum) in quarter_checks.items():
            if range_name not in wb.defined_names:
                print(f"FAIL: {cell_ref} - named range '{range_name}' missing, cannot verify sum")
                continue

            dn = wb.defined_names[range_name]
            # Parse the reference to extract cell range
            ref_text = dn.attr_text  # e.g., "Sheet1!$B$2:$B$4"
            # Extract the cell range part after '!'
            ref_part = ref_text.split('!')[-1] if '!' in ref_text else ref_text
            # Remove $ signs for easier parsing
            clean_ref = ref_part.replace('$', '')  # e.g., "B2:B4"

            # Parse start and end cells
            match = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', clean_ref)
            if not match:
                print(f"FAIL: {cell_ref} - cannot parse range '{ref_part}'")
                continue

            col_letter = match.group(1)
            start_row = int(match.group(2))
            end_row = int(match.group(4))

            # Sum the values in the range
            actual_sum = 0
            numeric_count = 0
            expected_count = end_row - start_row + 1
            for r in range(start_row, end_row + 1):
                cell_val = ws[f'{col_letter}{r}'].value
                if cell_val is not None and isinstance(cell_val, (int, float)):
                    actual_sum += cell_val
                    numeric_count += 1

            if numeric_count == expected_count and abs(actual_sum - expected_sum) < 0.01:
                print(f"PASS: {cell_ref} - SUM({range_name}) = {actual_sum} (expected {expected_sum})")
                value_score += 0.075
            else:
                print(f"FAIL: {cell_ref} - SUM({range_name}) = {actual_sum}, expected {expected_sum}")

        if value_score > 0:
            total_score += value_score
            print(f"  Component 3 subtotal: {value_score}/0.3")
    except Exception as e:
        print(f"ERROR: Component 3 (computed values) - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
