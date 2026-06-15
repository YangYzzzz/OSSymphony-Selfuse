"""
Initial Setup: ACL Best Paper Awards Affiliation Analysis
Task ID: osworld_multi_apps_acl_awards_calc_009
Domain: libreoffice_calc

Creates affiliation_analysis.ods with:
  - Column headers: Year, Paper Title, First Author, Affiliation Type (A1:D1)
  - Summary table headers in G1:H3 (Affiliation Type / Count, Academic, Industry)
  - NO data rows (the agent must fill these in from the ACL wiki)
  - Opens Chrome and LibreOffice Calc for the agent
"""

import os
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_acl_awards_calc_009'
XLSX_OUTPUT = f'{WORKDIR}/affiliation_analysis.xlsx'
OUTPUT = f'{WORKDIR}/affiliation_analysis.ods'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch a GUI app on the VM display without blocking script exit."""
    env = os.environ.copy()
    env['DISPLAY'] = ':0'
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # --- Main data table headers (A1:D1) ---
    headers = ['Year', 'Paper Title', 'First Author', 'Affiliation Type']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # --- Summary pivot-like table (G1:H3) ---
    # G1 = 'Affiliation Type', H1 = 'Count'
    ws.cell(row=1, column=7, value='Affiliation Type')
    ws.cell(row=1, column=8, value='Count')
    # G2 = 'Academic', G3 = 'Industry'  (H2, H3 are left empty — agent fills COUNTIF)
    ws.cell(row=2, column=7, value='Academic')
    ws.cell(row=3, column=7, value='Industry')
    # H2 and H3 intentionally left blank for agent to fill

    # NO data rows — agent must browse ACL wiki and add them

    # Save as xlsx first, then convert to ods via LibreOffice headless
    wb.save(XLSX_OUTPUT)
    print(f'XLSX file created: {XLSX_OUTPUT}')

    # Convert to ODS using LibreOffice headless
    env = os.environ.copy()
    env['DISPLAY'] = ':0'
    result = subprocess.run(
        ['libreoffice', '--headless', '--convert-to', 'ods', '--outdir', WORKDIR, XLSX_OUTPUT],
        env=env,
        capture_output=True,
        text=True,
        timeout=60
    )
    print(f'Conversion stdout: {result.stdout}')
    print(f'Conversion stderr: {result.stderr}')

    # Remove the intermediate xlsx
    if os.path.exists(XLSX_OUTPUT):
        os.remove(XLSX_OUTPUT)
        print(f'Removed intermediate XLSX')

    if os.path.exists(OUTPUT):
        print(f'ODS file created: {OUTPUT}')
    else:
        print(f'ERROR: ODS file not found at {OUTPUT}')

    # --- GUI-ready startup ---
    # Open Chrome first (agent needs to browse ACL wiki)
    launch_gui('google-chrome --new-window "https://aclweb.org/aclwiki/Best_paper_awards"',
               delay_sec=3.0)

    # Open LibreOffice Calc with the initial file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)

    print('GUI_READY: launched Chrome and LibreOffice Calc with DISPLAY=:0')


create_initial()
