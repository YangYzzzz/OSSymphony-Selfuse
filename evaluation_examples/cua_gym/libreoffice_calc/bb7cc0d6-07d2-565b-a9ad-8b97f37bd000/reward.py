"""
Reward Script: Marketing Campaign Performance Data Enhancement
Task ID: osworld_calc_multi_chart_computed_014
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Average row label in A10 = "Average"               (0.15 pts)
  Component 2: AVERAGE formulas in reach columns B10:G10          (0.20 pts)
  Component 3: AVERAGE formulas in conversion columns H10:M10     (0.15 pts)
  Component 4: Bar chart titled "Campaign Reach by Month" exists  (0.25 pts)
  Component 5: Line chart titled "Conversion Rate Trend" exists   (0.25 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_multi_chart_computed_014'


def get_chart_title(chart):
    """Extract title string from openpyxl chart title object."""
    try:
        title_obj = chart.title
        if title_obj is None:
            return None
        tx = title_obj.tx
        if tx and tx.rich and tx.rich.p:
            for para in tx.rich.p:
                if para.r:
                    for run in para.r:
                        if hasattr(run, 't') and run.t:
                            return run.t
        return None
    except Exception:
        return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Checks that:
    1. An average row has been added (row 10) with 'Average' in column A
    2. AVERAGE formulas cover the reach columns (B10:G10)
    3. AVERAGE formulas cover the conversion columns (H10:M10)
    4. A bar chart titled 'Campaign Reach by Month' has been created
    5. A line chart titled 'Conversion Rate Trend' has been created
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Campaign Performance' sheet must exist
    if 'Campaign Performance' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Campaign Performance' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Campaign Performance']

    # Component 1: Average row label — A10 should contain 'Average' (0.15 pts)
    try:
        a10_val = ws['A10'].value
        if a10_val is not None and str(a10_val).strip().lower() == 'average':
            print(f"PASS: Component 1 — A10 contains 'Average' label (value: {a10_val!r}) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — Expected 'Average' in A10, found: {a10_val!r}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: AVERAGE formulas for reach columns B10:G10 (0.20 pts)
    # The original reach data is in columns B-G (Jan-Jun Reach)
    try:
        reach_cols = ['B', 'C', 'D', 'E', 'F', 'G']
        reach_avg_formulas = 0
        for col in reach_cols:
            cell_val = ws[f'{col}10'].value
            if cell_val is not None and isinstance(cell_val, str) and 'AVERAGE' in cell_val.upper():
                reach_avg_formulas += 1
        if reach_avg_formulas == 6:
            print(f"PASS: Component 2 — All 6 reach AVERAGE formulas present in B10:G10 (0.20 pts)")
            total_score += 0.20
        elif reach_avg_formulas >= 3:
            print(f"PARTIAL: Component 2 — Only {reach_avg_formulas}/6 reach AVERAGE formulas found (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 — Only {reach_avg_formulas}/6 reach AVERAGE formulas in B10:G10")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: AVERAGE formulas for conversion columns H10:M10 (0.15 pts)
    # The original conversion data is in columns H-M (Jan-Jun Conv%)
    try:
        conv_cols = ['H', 'I', 'J', 'K', 'L', 'M']
        conv_avg_formulas = 0
        for col in conv_cols:
            cell_val = ws[f'{col}10'].value
            if cell_val is not None and isinstance(cell_val, str) and 'AVERAGE' in cell_val.upper():
                conv_avg_formulas += 1
        if conv_avg_formulas == 6:
            print(f"PASS: Component 3 — All 6 conversion AVERAGE formulas present in H10:M10 (0.15 pts)")
            total_score += 0.15
        elif conv_avg_formulas >= 3:
            print(f"PARTIAL: Component 3 — Only {conv_avg_formulas}/6 conversion AVERAGE formulas found (0.07 pts)")
            total_score += 0.07
        else:
            print(f"FAIL: Component 3 — Only {conv_avg_formulas}/6 conversion AVERAGE formulas in H10:M10")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Bar chart titled 'Campaign Reach by Month' (0.25 pts)
    try:
        charts = ws._charts
        bar_charts = [c for c in charts if type(c).__name__ == 'BarChart']
        bar_charts_with_title = [c for c in bar_charts
                                  if get_chart_title(c) is not None
                                  and 'Campaign Reach by Month'.lower() in get_chart_title(c).lower()]
        if len(bar_charts_with_title) >= 1:
            print(f"PASS: Component 4 — Bar chart titled 'Campaign Reach by Month' found (0.25 pts)")
            total_score += 0.25
        elif len(bar_charts) >= 1:
            titles = [get_chart_title(c) for c in bar_charts]
            print(f"PARTIAL: Component 4 — Bar chart found but title doesn't match 'Campaign Reach by Month', got: {titles} (0.12 pts)")
            total_score += 0.12
        else:
            print(f"FAIL: Component 4 — No bar chart found in sheet (total charts: {len(charts)})")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Line chart titled 'Conversion Rate Trend' (0.25 pts)
    try:
        charts = ws._charts
        line_charts = [c for c in charts if type(c).__name__ == 'LineChart']
        line_charts_with_title = [c for c in line_charts
                                   if get_chart_title(c) is not None
                                   and 'Conversion Rate Trend'.lower() in get_chart_title(c).lower()]
        if len(line_charts_with_title) >= 1:
            print(f"PASS: Component 5 — Line chart titled 'Conversion Rate Trend' found (0.25 pts)")
            total_score += 0.25
        elif len(line_charts) >= 1:
            titles = [get_chart_title(c) for c in line_charts]
            print(f"PARTIAL: Component 5 — Line chart found but title doesn't match 'Conversion Rate Trend', got: {titles} (0.12 pts)")
            total_score += 0.12
        else:
            print(f"FAIL: Component 5 — No line chart found in sheet (total charts: {len(charts)})")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the VM
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
