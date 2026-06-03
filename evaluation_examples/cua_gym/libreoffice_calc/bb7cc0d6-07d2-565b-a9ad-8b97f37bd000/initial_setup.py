"""
Initial Setup: Marketing Campaign Performance Spreadsheet
Task ID: osworld_calc_multi_chart_computed_014
Domain: libreoffice_calc

Creates the pre-task state: a spreadsheet with marketing campaign data
including campaign names, monthly reach figures (B-G) and conversion
rates (H-M) for 6 months. NO average rows and NO charts are present.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_multi_chart_computed_014'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Campaign Performance ---
    ws = wb.active
    ws.title = 'Campaign Performance'

    # Column headers
    headers = [
        'Campaign', 'Jan Reach', 'Feb Reach', 'Mar Reach', 'Apr Reach', 'May Reach', 'Jun Reach',
        'Jan Conv%', 'Feb Conv%', 'Mar Conv%', 'Apr Conv%', 'May Conv%', 'Jun Conv%'
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='FF1F497D', end_color='FF1F497D', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.row_dimensions[1].height = 30

    # Realistic marketing campaign data (5 campaigns × 6 months reach + 6 months conversion)
    # Reach values (number of people reached), Conversion % values
    campaigns = [
        # Campaign name, Jan-Jun Reach, Jan-Jun Conv%
        ('Spring Launch Email',    42300, 39800, 45600, 51200, 48700, 44100,  3.2, 3.5, 3.8, 4.1, 3.9, 3.6),
        ('Social Media Blitz',     87500, 92400, 105300, 118200, 124600, 132000,  2.1, 2.4, 2.7, 3.0, 3.2, 3.5),
        ('Influencer Partnership', 63200, 71400, 78900, 82500, 79300, 74600,  4.5, 4.8, 5.2, 5.0, 4.7, 4.4),
        ('Search Engine Ads',      55100, 57800, 61400, 65900, 70200, 74500,  5.8, 6.1, 6.3, 6.5, 6.7, 6.9),
        ('Content Marketing Hub',  28400, 31200, 35700, 41100, 46500, 52300,  7.2, 7.5, 7.8, 8.1, 8.4, 8.7),
        ('Podcast Sponsorships',   19800, 22300, 25400, 28600, 31900, 35200,  9.1, 9.3, 9.6, 9.8, 10.1, 10.4),
        ('Retargeting Display',    34700, 36500, 39800, 42100, 44600, 46900,  6.3, 6.5, 6.8, 7.1, 7.3, 7.6),
        ('Affiliate Network',      21600, 23900, 27100, 30400, 33800, 37200,  4.1, 4.3, 4.6, 4.9, 5.1, 5.4),
    ]

    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r, row_data in enumerate(campaigns, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            if c == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='right', vertical='center')
            # Format reach values as integers, conv% as percentage
            if 2 <= c <= 7:
                cell.number_format = '#,##0'
            elif 8 <= c <= 13:
                cell.number_format = '0.0'

    # Column widths
    ws.column_dimensions['A'].width = 24
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].width = 12
    for col_letter in ['H', 'I', 'J', 'K', 'L', 'M']:
        ws.column_dimensions[col_letter].width = 11

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open LibreOffice Calc with the initial file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
