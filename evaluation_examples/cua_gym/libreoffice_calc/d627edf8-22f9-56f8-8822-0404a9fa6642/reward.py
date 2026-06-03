"""
Reward Script: Apply Auto Outline feature to spreadsheet with summary formulas
Task ID: calc_adv_group_autooutline_040
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Row groups for North region detail rows (rows 2-5) at outline_level=1
  Component 2 (0.3): Row groups for South (rows 7-11) and East (rows 13-16) detail rows at outline_level=1
  Component 3 (0.3): Column groups for quarterly columns (C,D at level 2; E,H at level 1; F,G at level 2)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_adv_group_autooutline_040'
SHEET_NAME = 'Summary Report'


def verify_task(file_path):
    """
    Verify that Auto Outline was applied correctly to the spreadsheet.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: verify the sheet exists
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found in file. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Component 1: Row groups for North region detail rows (rows 2-5) at outline_level=1 (0.4 points)
    # AutoOutline detects North detail rows (2-5) under North Total (row 6), assigns outline_level=1
    # Initial file has NO outline levels set — this check only passes on the golden file
    try:
        north_detail_rows = [2, 3, 4, 5]
        north_matched = sum(
            1 for row_idx in north_detail_rows
            if ws.row_dimensions.get(row_idx) is not None
            and ws.row_dimensions[row_idx].outline_level == 1
        )
        if north_matched == len(north_detail_rows):
            print(f"PASS: Component 1 — All North detail rows (2-5) have outline_level=1 (0.4 pts)")
            print(f"  All {north_matched}/{len(north_detail_rows)} rows confirmed at level 1")
            total_score += 0.4
        else:
            missing = [r for r in north_detail_rows
                       if ws.row_dimensions.get(r) is None
                       or ws.row_dimensions[r].outline_level != 1]
            print(f"FAIL: Component 1 — Only {north_matched}/{len(north_detail_rows)} North rows at outline_level=1")
            print(f"  Missing/incorrect rows: {missing}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Row groups for South (rows 7-11) and East (rows 13-16) at outline_level=1 (0.3 points)
    # AutoOutline detects South and East detail rows, assigning outline_level=1 for each region
    try:
        south_detail_rows = [7, 8, 9, 10, 11]
        east_detail_rows = [13, 14, 15, 16]
        all_region_rows = south_detail_rows + east_detail_rows

        region_matched = sum(
            1 for row_idx in all_region_rows
            if ws.row_dimensions.get(row_idx) is not None
            and ws.row_dimensions[row_idx].outline_level == 1
        )
        if region_matched == len(all_region_rows):
            print(f"PASS: Component 2 — All South (7-11) and East (13-16) detail rows at outline_level=1 (0.3 pts)")
            print(f"  All {region_matched}/{len(all_region_rows)} rows confirmed at level 1")
            total_score += 0.3
        else:
            missing = [r for r in all_region_rows
                       if ws.row_dimensions.get(r) is None
                       or ws.row_dimensions[r].outline_level != 1]
            print(f"FAIL: Component 2 — Only {region_matched}/{len(all_region_rows)} South/East rows at outline_level=1")
            print(f"  Missing/incorrect rows: {missing}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Column groups for quarterly columns (0.3 points)
    # AutoOutline groups: C,D (quarterly Q1/Q2) under H1 Total (E) at level 2;
    # E (H1 Total) at level 1; F,G (quarterly Q3/Q4) under H2 Total (H) at level 2;
    # H (H2 Total) at level 1
    # Initial file has NO column outline levels — all pass only on golden file
    try:
        expected_col_levels = {
            'C': 2,
            'D': 2,
            'E': 1,
            'F': 2,
            'G': 2,
            'H': 1,
        }
        col_matched = sum(
            1 for col_letter, expected_level in expected_col_levels.items()
            if ws.column_dimensions.get(col_letter) is not None
            and ws.column_dimensions[col_letter].outline_level == expected_level
        )
        if col_matched == len(expected_col_levels):
            print(f"PASS: Component 3 — All column groups correctly structured (0.3 pts)")
            details = [f"Col {c}: level={ws.column_dimensions[c].outline_level}" for c in expected_col_levels]
            print(f"  Details: {details}")
            total_score += 0.3
        else:
            wrong_cols = [
                f"Col {c}: expected={lvl}, actual={ws.column_dimensions[c].outline_level if ws.column_dimensions.get(c) else 'not set'}"
                for c, lvl in expected_col_levels.items()
                if ws.column_dimensions.get(c) is None or ws.column_dimensions[c].outline_level != lvl
            ]
            print(f"FAIL: Component 3 — Only {col_matched}/{len(expected_col_levels)} columns at correct outline levels")
            print(f"  Incorrect: {wrong_cols}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
