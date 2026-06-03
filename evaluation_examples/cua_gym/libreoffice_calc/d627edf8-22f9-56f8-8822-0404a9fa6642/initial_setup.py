"""
Initial Setup: Apply Auto Outline feature to a spreadsheet with summary formulas
Task ID: calc_adv_group_autooutline_040
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_adv_group_autooutline_040'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Summary Report ---
    ws = wb.active
    ws.title = 'Summary Report'

    # Column headers
    # A: Region, B: Product, C: Q1, D: Q2, E: Q1+Q2 Subtotal, F: Q3, G: Q4, H: Q3+Q4 Subtotal, I: Annual Total
    headers = ['Region', 'Product', 'Q1 Sales', 'Q2 Sales', 'H1 Total', 'Q3 Sales', 'Q4 Sales', 'H2 Total', 'Annual Total']
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Row structure (matching task context):
    # Rows 2-5: Region North detail rows
    # Row 6: North Total
    # Rows 7-11: Region South detail rows
    # Row 12: South Total
    # Rows 13-16: Region East detail rows
    # Row 17: East Total
    # Row 18: Grand Total

    # North detail rows (2-5)
    north_data = [
        ['North', 'Electronics',   45230, 52100, None, 48750, 51200, None, None],
        ['North', 'Furniture',     28400, 31500, None, 29800, 33200, None, None],
        ['North', 'Appliances',    19800, 22300, None, 21400, 24100, None, None],
        ['North', 'Office Supplies', 8700, 9800,  None, 9200,  10100, None, None],
    ]

    # South detail rows (7-11)
    south_data = [
        ['South', 'Electronics',   61200, 68400, None, 63500, 71200, None, None],
        ['South', 'Furniture',     38500, 42700, None, 40100, 44800, None, None],
        ['South', 'Appliances',    27600, 30200, None, 28900, 32400, None, None],
        ['South', 'Office Supplies', 11200, 12800, None, 11900, 13600, None, None],
        ['South', 'Accessories',  15400, 17100, None, 16200, 18300, None, None],
    ]

    # East detail rows (13-16)
    east_data = [
        ['East', 'Electronics',   38700, 43500, None, 40200, 45900, None, None],
        ['East', 'Furniture',     24100, 27300, None, 25600, 29100, None, None],
        ['East', 'Appliances',    16900, 18700, None, 17800, 20300, None, None],
        ['East', 'Office Supplies', 7300, 8200, None, 7900,  9000,  None, None],
    ]

    # Write detail rows
    for r, row_data in enumerate(north_data, 2):
        for c, val in enumerate(row_data, 1):
            if val is not None:
                ws.cell(row=r, column=c, value=val)

    # Row 6: North Total
    total_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    total_font = Font(bold=True)
    row6 = ws.row_dimensions[6]
    ws.cell(row=6, column=1, value='North Total')
    ws.cell(row=6, column=2, value='All Products')
    ws.cell(row=6, column=3, value='=SUM(C2:C5)')
    ws.cell(row=6, column=4, value='=SUM(D2:D5)')
    ws.cell(row=6, column=5, value='=SUM(E2:E5)')
    ws.cell(row=6, column=6, value='=SUM(F2:F5)')
    ws.cell(row=6, column=7, value='=SUM(G2:G5)')
    ws.cell(row=6, column=8, value='=SUM(H2:H5)')
    ws.cell(row=6, column=9, value='=SUM(I2:I5)')
    for c in range(1, 10):
        ws.cell(row=6, column=c).fill = total_fill
        ws.cell(row=6, column=c).font = total_font

    # South detail rows (7-11)
    for r, row_data in enumerate(south_data, 7):
        for c, val in enumerate(row_data, 1):
            if val is not None:
                ws.cell(row=r, column=c, value=val)

    # Row 12: South Total
    ws.cell(row=12, column=1, value='South Total')
    ws.cell(row=12, column=2, value='All Products')
    ws.cell(row=12, column=3, value='=SUM(C7:C11)')
    ws.cell(row=12, column=4, value='=SUM(D7:D11)')
    ws.cell(row=12, column=5, value='=SUM(E7:E11)')
    ws.cell(row=12, column=6, value='=SUM(F7:F11)')
    ws.cell(row=12, column=7, value='=SUM(G7:G11)')
    ws.cell(row=12, column=8, value='=SUM(H7:H11)')
    ws.cell(row=12, column=9, value='=SUM(I7:I11)')
    for c in range(1, 10):
        ws.cell(row=12, column=c).fill = total_fill
        ws.cell(row=12, column=c).font = total_font

    # East detail rows (13-16)
    for r, row_data in enumerate(east_data, 13):
        for c, val in enumerate(row_data, 1):
            if val is not None:
                ws.cell(row=r, column=c, value=val)

    # Row 17: East Total
    ws.cell(row=17, column=1, value='East Total')
    ws.cell(row=17, column=2, value='All Products')
    ws.cell(row=17, column=3, value='=SUM(C13:C16)')
    ws.cell(row=17, column=4, value='=SUM(D13:D16)')
    ws.cell(row=17, column=5, value='=SUM(E13:E16)')
    ws.cell(row=17, column=6, value='=SUM(F13:F16)')
    ws.cell(row=17, column=7, value='=SUM(G13:G16)')
    ws.cell(row=17, column=8, value='=SUM(H13:H16)')
    ws.cell(row=17, column=9, value='=SUM(I13:I16)')
    for c in range(1, 10):
        ws.cell(row=17, column=c).fill = total_fill
        ws.cell(row=17, column=c).font = total_font

    # Row 18: Grand Total
    grand_fill = PatternFill(start_color='FF1F4E79', end_color='FF1F4E79', fill_type='solid')
    grand_font = Font(bold=True, color='FFFFFFFF')
    ws.cell(row=18, column=1, value='Grand Total')
    ws.cell(row=18, column=2, value='All Regions')
    ws.cell(row=18, column=3, value='=SUM(C6,C12,C17)')
    ws.cell(row=18, column=4, value='=SUM(D6,D12,D17)')
    ws.cell(row=18, column=5, value='=SUM(E6,E12,E17)')
    ws.cell(row=18, column=6, value='=SUM(F6,F12,F17)')
    ws.cell(row=18, column=7, value='=SUM(G6,G12,G17)')
    ws.cell(row=18, column=8, value='=SUM(H6,H12,H17)')
    ws.cell(row=18, column=9, value='=SUM(I6,I12,I17)')
    for c in range(1, 10):
        ws.cell(row=18, column=c).fill = grand_fill
        ws.cell(row=18, column=c).font = grand_font

    # Add H1 Total formula (column E) for detail rows
    # E = C + D for each detail row
    for r in list(range(2, 6)) + list(range(7, 12)) + list(range(13, 17)):
        ws.cell(row=r, column=5, value=f'=C{r}+D{r}')

    # Add H2 Total formula (column H) for detail rows
    # H = F + G for each detail row
    for r in list(range(2, 6)) + list(range(7, 12)) + list(range(13, 17)):
        ws.cell(row=r, column=8, value=f'=F{r}+G{r}')

    # Add Annual Total formula (column I) for detail rows
    # I = E + H for each detail row
    for r in list(range(2, 6)) + list(range(7, 12)) + list(range(13, 17)):
        ws.cell(row=r, column=9, value=f'=E{r}+H{r}')

    # Column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 18
    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 12

    # Freeze header row
    ws.freeze_panes = 'A2'

    # IMPORTANT: No outline groups set in initial file
    # The task is to APPLY Auto Outline - so no groups should exist yet

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: Summary Report')
    print('Rows: 1 header + 4 North detail + 1 North Total + 5 South detail + 1 South Total + 4 East detail + 1 East Total + 1 Grand Total = 18 rows')
    print('No outline groups set (task is to apply Auto Outline)')


create_initial()
