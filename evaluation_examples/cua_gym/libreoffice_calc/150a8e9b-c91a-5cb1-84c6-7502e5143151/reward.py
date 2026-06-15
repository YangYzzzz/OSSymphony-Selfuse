"""
Reward Script: Sort sales data by Region/Date, add subtotals, outline grouping, print setup, header/footer
Task ID: calc_wf_005
Domain: libreoffice_calc
Scoring:
  Component 1: Region sort (ascending) — 0.20
  Component 2: Date sort within region (ascending) — 0.15
  Component 3: SUBTOTAL formulas per region — 0.25
  Component 4: Outline grouping levels — 0.15
  Component 5: Print area + repeating title rows — 0.10
  Component 6: Header 'Regional Sales Report' + footer with page number — 0.15
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_005'


def persist_app_state(domain: str):
    """Best-effort save via Ctrl+S for unsaved GUI edits."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Sales' not in wb.sheetnames:
        print("CRITICAL: 'Sales' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Sales']

    # =========================================================================
    # Collect data rows vs subtotal rows
    # We need to identify which rows are data and which are subtotal/grand total
    # =========================================================================
    max_row = ws.max_row
    data_rows = []       # (row_num, region, date_val, revenue)
    subtotal_rows = []   # (row_num, label, formula)
    grand_total_row = None

    for r in range(2, max_row + 1):
        b_val = ws.cell(r, 2).value  # Region column
        e_val = ws.cell(r, 5).value  # Revenue column

        if b_val is not None and isinstance(b_val, str) and b_val.strip().endswith('Total'):
            if b_val.strip() == 'Grand Total':
                grand_total_row = (r, b_val.strip(), e_val)
            else:
                subtotal_rows.append((r, b_val.strip(), e_val))
        elif b_val is not None:
            data_rows.append((r, str(b_val).strip(), ws.cell(r, 1).value, e_val))

    # =========================================================================
    # Component 1: Data sorted by Region ascending (0.20 points)
    # The regions across data rows should appear in ascending alphabetical order
    # (grouped: all East, then all North, then all South, then all West)
    # =========================================================================
    try:
        if len(data_rows) < 10:
            print(f"FAIL: Component 1 — Too few data rows ({len(data_rows)}), expected ~60")
        else:
            regions_in_order = [row[1] for row in data_rows]
            # Check that regions are grouped and in ascending order
            seen_regions = []
            current_region = None
            region_sort_ok = (len(regions_in_order) > 0)  # start optimistic, falsify below
            for reg in regions_in_order:
                if reg != current_region:
                    if reg in seen_regions:
                        # Region appeared before, broke grouping
                        region_sort_ok = False
                        break
                    seen_regions.append(reg)
                    current_region = reg

            if region_sort_ok and seen_regions == sorted(seen_regions):
                print(f"PASS: Component 1 — Regions sorted ascending: {seen_regions} (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 1 — Regions not sorted ascending. Order found: {seen_regions}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Dates sorted ascending within each region (0.15 points)
    # Within each region group, dates should be in ascending order
    # =========================================================================
    try:
        if len(data_rows) < 10:
            print(f"FAIL: Component 2 — Too few data rows")
        else:
            # Group data rows by region
            from collections import OrderedDict
            region_groups = OrderedDict()
            for row_num, region, date_val, rev in data_rows:
                if region not in region_groups:
                    region_groups[region] = []
                region_groups[region].append(date_val)

            dates_sort_ok = (len(region_groups) > 0)  # start optimistic, falsify below
            failed_region = None
            for region, dates in region_groups.items():
                # Convert to comparable format; dates could be datetime or string
                comparable_dates = []
                for d in dates:
                    if d is not None:
                        comparable_dates.append(str(d))
                    else:
                        comparable_dates.append('')

                if comparable_dates != sorted(comparable_dates):
                    dates_sort_ok = False
                    failed_region = region
                    break

            if dates_sort_ok and len(region_groups) >= 2:
                print(f"PASS: Component 2 — Dates sorted ascending within each region ({len(region_groups)} regions) (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 2 — Dates not sorted within region '{failed_region}'")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: SUBTOTAL formulas for each region + grand total (0.25 points)
    # There should be subtotal rows with SUBTOTAL(9,...) formulas in column E
    # for each region, plus a grand total row.
    # =========================================================================
    try:
        subtotal_count = 0
        grand_total_found = (grand_total_row is not None)  # derive from actual data

        for row_num, label, formula in subtotal_rows:
            if isinstance(formula, str) and 'SUBTOTAL' in formula.upper():
                subtotal_count += 1

        has_gt_formula = False
        if grand_total_found:
            row_num, label, formula = grand_total_row
            if isinstance(formula, str) and 'SUBTOTAL' in formula.upper():
                has_gt_formula = (len(formula) > 0)  # derived from formula check

        # We expect 4 region subtotals + 1 grand total
        if subtotal_count >= 4 and has_gt_formula:
            print(f"PASS: Component 3 — {subtotal_count} region subtotals + grand total with SUBTOTAL formulas (0.25 pts)")
            total_score += 0.25
        elif subtotal_count >= 2:
            partial = 0.10
            print(f"PARTIAL: Component 3 — Only {subtotal_count} subtotals found, grand_total={has_gt_formula} ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Found {subtotal_count} region subtotals, grand_total={has_gt_formula}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: Outline grouping levels present (0.15 points)
    # Data rows should have outline_level > 0 (level 2 for data, level 1 for subtotals)
    # =========================================================================
    try:
        rows_with_outline = 0
        rows_level_1 = 0
        rows_level_2 = 0
        for r in range(2, max_row + 1):
            rd = ws.row_dimensions.get(r)
            if rd and rd.outline_level and rd.outline_level > 0:
                rows_with_outline += 1
                if rd.outline_level == 1:
                    rows_level_1 += 1
                elif rd.outline_level == 2:
                    rows_level_2 += 1

        if rows_with_outline >= 10 and rows_level_1 >= 2 and rows_level_2 >= 10:
            print(f"PASS: Component 4 — Outline grouping: {rows_level_2} rows at level 2, {rows_level_1} rows at level 1 (0.15 pts)")
            total_score += 0.15
        elif rows_with_outline >= 5:
            partial = 0.07
            print(f"PARTIAL: Component 4 — Some outline grouping found ({rows_with_outline} rows) but incomplete ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — No significant outline grouping found ({rows_with_outline} rows with outline)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # =========================================================================
    # Component 5: Print area set + row 1 as repeating title row (0.10 points)
    # =========================================================================
    try:
        print_area = ws.print_area
        print_title_rows = ws.print_title_rows

        pa_ok = False
        ptr_ok = False

        if print_area:
            # print_area could be a string like "'Sales'!$A$1:$E$66" or a list
            pa_str = str(print_area)
            pa_ok = ('A' in pa_str and 'E' in pa_str)
            if pa_ok:
                print(f"  Print area found: {pa_str}")

        if print_title_rows:
            ptr_str = str(print_title_rows)
            ptr_ok = ('1' in ptr_str)
            if ptr_ok:
                print(f"  Print title rows found: {ptr_str}")

        if pa_ok and ptr_ok:
            print(f"PASS: Component 5 — Print area and repeating title rows set (0.10 pts)")
            total_score += 0.10
        elif pa_ok or ptr_ok:
            partial = 0.05
            print(f"PARTIAL: Component 5 — print_area={pa_ok}, print_title_rows={ptr_ok} ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — No print area or title rows (print_area={print_area}, title_rows={print_title_rows})")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # =========================================================================
    # Component 6: Header 'Regional Sales Report' + footer with page number (0.15 points)
    # =========================================================================
    try:
        header_text = ws.oddHeader.center.text if ws.oddHeader and ws.oddHeader.center else ''
        if not header_text:
            # Try raw string
            header_raw = str(ws.oddHeader) if ws.oddHeader else ''
            header_text = header_raw

        footer_text = ws.oddFooter.center.text if ws.oddFooter and ws.oddFooter.center else ''
        if not footer_text:
            footer_raw = str(ws.oddFooter) if ws.oddFooter else ''
            footer_text = footer_raw

        header_ok = 'Regional Sales Report' in str(header_text)
        footer_ok = '&P' in str(footer_text) or 'Page' in str(footer_text) or '&p' in str(footer_text)

        if header_ok and footer_ok:
            print(f"PASS: Component 6 — Header='{header_text}', Footer='{footer_text}' (0.15 pts)")
            total_score += 0.15
        elif header_ok or footer_ok:
            partial = 0.07
            print(f"PARTIAL: Component 6 — header_ok={header_ok}, footer_ok={footer_ok} ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 6 — header='{header_text}', footer='{footer_text}'")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
