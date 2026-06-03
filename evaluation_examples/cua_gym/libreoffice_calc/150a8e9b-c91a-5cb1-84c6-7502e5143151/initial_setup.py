"""
Initial Setup: Sort sales data, add subtotals, outline grouping, print area with headers/footers
Task ID: calc_wf_005
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random
from datetime import date, timedelta
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_005'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sales'

    # Headers
    headers = ['Date', 'Region', 'Salesperson', 'Product', 'Revenue']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic data pools
    regions = ['North', 'South', 'East', 'West']
    salespersons = {
        'North': ['Sarah Chen', 'Marcus Johnson', 'Emily Park', 'David Kim'],
        'South': ['Maria Garcia', 'James Wilson', 'Priya Patel', 'Carlos Rivera'],
        'East': ['Lisa Wang', 'Robert Taylor', 'Aisha Khan', 'Tom Anderson'],
        'West': ['Rachel Green', 'Mike Torres', 'Jennifer Lee', 'Brandon Scott'],
    }
    products = [
        'Enterprise License', 'Cloud Storage Plan', 'Security Suite',
        'Analytics Dashboard', 'API Gateway', 'Backup Solution',
        'Collaboration Tools', 'DevOps Platform', 'Data Warehouse',
        'Mobile SDK',
    ]

    # Generate 60 rows of data across 4 regions, dates Jan-Jun 2024
    # Deliberately NOT sorted - the task requires the agent to sort
    random.seed(42)
    start_date = date(2024, 1, 1)
    end_date = date(2024, 6, 30)
    date_range_days = (end_date - start_date).days

    data = []
    for i in range(60):
        region = regions[i % 4]  # distribute evenly but interleaved
        d = start_date + timedelta(days=random.randint(0, date_range_days))
        sp = random.choice(salespersons[region])
        prod = random.choice(products)
        revenue = round(random.uniform(1500, 45000), 2)
        data.append([d, region, sp, prod, revenue])

    # Shuffle to ensure unsorted state
    random.shuffle(data)

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        ws.cell(row=r, column=1).number_format = 'yyyy-mm-dd'
        ws.cell(row=r, column=2, value=row_data[1])
        ws.cell(row=r, column=3, value=row_data[2])
        ws.cell(row=r, column=4, value=row_data[3])
        ws.cell(row=r, column=5, value=row_data[4])
        ws.cell(row=r, column=5).number_format = '$#,##0.00'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 14

    # NO sorting, NO subtotals, NO outline, NO print area, NO headers/footers
    # The agent must do all of that

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
