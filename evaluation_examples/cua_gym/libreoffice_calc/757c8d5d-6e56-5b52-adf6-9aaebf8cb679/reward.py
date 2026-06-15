"""
Reward Script: Filter pivot table to show only East and West regions
Task ID: calc_adv_pivot_filter_005
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.4): 'North' region row is absent from the Report pivot table
  - Component 2 (0.4): 'South' region row is absent from the Report pivot table
  - Component 3 (0.2): Grand Total row reflects only East+West data (B col = 91500)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_adv_pivot_filter_005'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    The task requires filtering the pivot table on sheet 'Report' to show
    only East and West regions — removing North and South from the view.
    """
    total_score = 0.0

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: Report sheet must exist
    if 'Report' not in wb.sheetnames:
        print("CRITICAL: 'Report' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Report']

    # Collect all non-None values in column A (Region column)
    region_values = []
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is not None:
            region_values.append(str(val).strip())

    print(f"INFO: Region column values found: {region_values}")

    # Component 1: 'North' region is absent from the pivot table (0.4 points)
    # The task asks to remove North from the view; initial file has 'North' in row 3
    try:
        north_absent = 'North' not in region_values
        if north_absent:
            print(f"PASS: Component 1 — 'North' region is NOT present in Report sheet (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — 'North' region still present in Report sheet; expected it removed")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: 'South' region is absent from the pivot table (0.4 points)
    # The task asks to remove South from the view; initial file has 'South' in row 4
    try:
        south_absent = 'South' not in region_values
        if south_absent:
            print(f"PASS: Component 2 — 'South' region is NOT present in Report sheet (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 2 — 'South' region still present in Report sheet; expected it removed")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Grand Total row reflects only East+West aggregation (0.2 points)
    # When only East and West remain, the Grand Total for Gadget (col B) should be 91500
    # (East=52000, West=39500). Initial file had Grand Total Gadget = 146150 (all 4 regions).
    try:
        expected_gadget_total = 91500
        # Find the Grand Total row
        grand_total_row = None
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val is not None and str(val).strip() == 'Grand Total':
                grand_total_row = row
                break

        if grand_total_row is None:
            print(f"FAIL: Component 3 — 'Grand Total' row not found in Report sheet")
        else:
            gadget_total = ws.cell(row=grand_total_row, column=2).value
            try:
                gadget_total_num = float(gadget_total)
            except (TypeError, ValueError):
                gadget_total_num = None

            if gadget_total_num is not None and abs(gadget_total_num - expected_gadget_total) < 0.01:
                print(f"PASS: Component 3 — Grand Total Gadget = {gadget_total_num} (expected {expected_gadget_total}, East+West only) (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 3 — Grand Total Gadget = {gadget_total_num}, expected {expected_gadget_total} (East+West only)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
