"""
Initial Setup: Pivot table showing Sum of Revenue by Region and Product (all 4 regions)
Task ID: calc_adv_pivot_filter_005
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_adv_pivot_filter_005'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    # ─── Sheet 1: RawSales (source data) ───────────────────────────────────────
    ws_raw = wb.active
    ws_raw.title = 'RawSales'

    raw_headers = ['OrderID', 'Region', 'Product', 'SalesRep', 'Revenue', 'Date']
    for col, h in enumerate(raw_headers, 1):
        cell = ws_raw.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFFFF')

    raw_data = [
        [1001, 'East',  'Gadget',  'Alice Kim',     12500, '2024-01-05'],
        [1002, 'West',  'Widget',  'Brian Torres',  18400, '2024-01-08'],
        [1003, 'North', 'Gadget',  'Carol White',    9800, '2024-01-10'],
        [1004, 'South', 'Widget',  'David Green',   11200, '2024-01-12'],
        [1005, 'East',  'Widget',  'Alice Kim',     15600, '2024-01-15'],
        [1006, 'West',  'Gadget',  'Fiona Hart',    22300, '2024-01-18'],
        [1007, 'North', 'Sprocket','George Yuen',    7400, '2024-01-20'],
        [1008, 'South', 'Sprocket','Hana Patel',     6900, '2024-01-22'],
        [1009, 'East',  'Sprocket','Alice Kim',     10200, '2024-02-03'],
        [1010, 'West',  'Sprocket','Brian Torres',  14100, '2024-02-07'],
        [1011, 'North', 'Widget',  'Carol White',   13600, '2024-02-09'],
        [1012, 'South', 'Gadget',  'David Green',    8750, '2024-02-11'],
        [1013, 'East',  'Gadget',  'Elena Morris',  19400, '2024-02-14'],
        [1014, 'West',  'Widget',  'Fiona Hart',    21500, '2024-02-16'],
        [1015, 'North', 'Gadget',  'George Yuen',   11300, '2024-02-19'],
        [1016, 'South', 'Widget',  'Hana Patel',    16800, '2024-02-21'],
        [1017, 'East',  'Widget',  'Alice Kim',     13900, '2024-03-02'],
        [1018, 'West',  'Sprocket','Brian Torres',   9600, '2024-03-05'],
        [1019, 'North', 'Sprocket','Ivan Reyes',    12400, '2024-03-08'],
        [1020, 'South', 'Sprocket','Hana Patel',     8100, '2024-03-10'],
        [1021, 'East',  'Sprocket','Elena Morris',  11700, '2024-03-13'],
        [1022, 'West',  'Gadget',  'Fiona Hart',    17200, '2024-03-15'],
        [1023, 'North', 'Widget',  'Ivan Reyes',     9500, '2024-03-18'],
        [1024, 'South', 'Gadget',  'David Green',   14600, '2024-03-20'],
        [1025, 'East',  'Gadget',  'Alice Kim',     20100, '2024-03-22'],
        [1026, 'West',  'Widget',  'Brian Torres',  16700, '2024-03-25'],
        [1027, 'North', 'Gadget',  'Carol White',   10200, '2024-03-27'],
        [1028, 'South', 'Widget',  'Hana Patel',    12300, '2024-03-29'],
    ]

    for r, row_data in enumerate(raw_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_raw.cell(row=r, column=c, value=val)

    ws_raw.column_dimensions['A'].width = 10
    ws_raw.column_dimensions['B'].width = 10
    ws_raw.column_dimensions['C'].width = 12
    ws_raw.column_dimensions['D'].width = 18
    ws_raw.column_dimensions['E'].width = 12
    ws_raw.column_dimensions['F'].width = 14

    # ─── Sheet 2: Report (pivot table result — all 4 regions) ─────────────────
    ws_report = wb.create_sheet('Report')

    # Pivot table header area
    # Layout: A1:F20 approx
    # Row 1: "Sum of Revenue" label + Product column headers
    # Rows 2-5: Region rows (East, North, South, West)
    # Row 6: Grand Total

    # Products (columns)
    products = ['Gadget', 'Sprocket', 'Widget']
    # Region totals (pre-calculated from raw_data above)
    # East:    Gadget=12500+19400+20100=51900  Sprocket=10200+11700=21900  Widget=15600+13900=29500  Total=103400 -> wait recalc
    # Let me compute properly:
    # East-Gadget:   1001(12500)+1013(19400)+1025(20100) = 52000
    # East-Widget:   1005(15600)+1017(13900) = 29500
    # East-Sprocket: 1009(10200)+1021(11700) = 21900
    # East Total: 52000+29500+21900 = 103400

    # West-Gadget:   1006(22300)+1022(17200) = 39500
    # West-Widget:   1002(18400)+1014(21500)+1026(16700) = 56600
    # West-Sprocket: 1010(14100)+1018(9600) = 23700
    # West Total: 39500+56600+23700 = 119800

    # North-Gadget:   1003(9800)+1015(11300)+1027(10200) = 31300
    # North-Widget:   1011(13600)+1023(9500) = 23100
    # North-Sprocket: 1007(7400)+1019(12400) = 19800
    # North Total: 31300+23100+19800 = 74200

    # South-Gadget:   1012(8750)+1024(14600) = 23350
    # South-Widget:   1004(11200)+1016(16800)+1028(12300) = 40300
    # South-Sprocket: 1008(6900)+1020(8100) = 15000
    # South Total: 23350+40300+15000 = 78650

    # Grand Totals per product:
    # Gadget:   52000+39500+31300+23350 = 146150
    # Widget:   29500+56600+23100+40300 = 149500
    # Sprocket: 21900+23700+19800+15000 = 80400
    # Grand Total: 103400+119800+74200+78650 = 376050

    pivot_data = {
        'East':  {'Gadget': 52000, 'Sprocket': 21900, 'Widget': 29500, 'Total': 103400},
        'North': {'Gadget': 31300, 'Sprocket': 19800, 'Widget': 23100, 'Total': 74200},
        'South': {'Gadget': 23350, 'Sprocket': 15000, 'Widget': 40300, 'Total': 78650},
        'West':  {'Gadget': 39500, 'Sprocket': 23700, 'Widget': 56600, 'Total': 119800},
    }

    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF')
    total_fill  = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    total_font  = Font(bold=True)
    thin_side   = Side(style='thin', color='FF000000')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Row 1: headers
    ws_report.cell(row=1, column=1, value='Region').font = header_font
    ws_report['A1'].fill = header_fill
    ws_report['A1'].border = thin_border

    for col_idx, product in enumerate(products, 2):
        cell = ws_report.cell(row=1, column=col_idx, value=product)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    ws_report.cell(row=1, column=5, value='Grand Total').font = header_font
    ws_report['E1'].fill = header_fill
    ws_report['E1'].alignment = Alignment(horizontal='center')
    ws_report['E1'].border = thin_border

    # Rows 2-5: regions (East, North, South, West — sorted alphabetically)
    regions_sorted = ['East', 'North', 'South', 'West']
    for row_idx, region in enumerate(regions_sorted, 2):
        d = pivot_data[region]
        ws_report.cell(row=row_idx, column=1, value=region).border = thin_border
        ws_report.cell(row=row_idx, column=1).font = Font(bold=False)

        ws_report.cell(row=row_idx, column=2, value=d['Gadget']).border = thin_border
        ws_report.cell(row=row_idx, column=2).number_format = '#,##0'
        ws_report.cell(row=row_idx, column=3, value=d['Sprocket']).border = thin_border
        ws_report.cell(row=row_idx, column=3).number_format = '#,##0'
        ws_report.cell(row=row_idx, column=4, value=d['Widget']).border = thin_border
        ws_report.cell(row=row_idx, column=4).number_format = '#,##0'
        ws_report.cell(row=row_idx, column=5, value=d['Total']).border = thin_border
        ws_report.cell(row=row_idx, column=5).number_format = '#,##0'
        ws_report.cell(row=row_idx, column=5).font = Font(bold=True)

    # Row 6: Grand Total row
    gt_row = 6
    ws_report.cell(row=gt_row, column=1, value='Grand Total').font = total_font
    ws_report['A6'].fill = total_fill
    ws_report['A6'].border = thin_border

    gt_gadget   = sum(pivot_data[r]['Gadget']   for r in regions_sorted)
    gt_sprocket = sum(pivot_data[r]['Sprocket'] for r in regions_sorted)
    gt_widget   = sum(pivot_data[r]['Widget']   for r in regions_sorted)
    gt_total    = sum(pivot_data[r]['Total']    for r in regions_sorted)

    for col_idx, val in enumerate([gt_gadget, gt_sprocket, gt_widget, gt_total], 2):
        cell = ws_report.cell(row=gt_row, column=col_idx, value=val)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thin_border
        cell.number_format = '#,##0'

    ws_report.column_dimensions['A'].width = 15
    ws_report.column_dimensions['B'].width = 14
    ws_report.column_dimensions['C'].width = 14
    ws_report.column_dimensions['D'].width = 14
    ws_report.column_dimensions['E'].width = 14

    ws_report.freeze_panes = 'B2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('  Sheets: RawSales (source data), Report (pivot table — all 4 regions)')
    print('  Report: East, North, South, West all visible')

create_initial()
