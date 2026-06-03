"""
FINAL REWARD SCRIPT - SUCCESS
Task: Copy all entries from 'Transaction ID' to 'Padded Transaction' and prepend zeros to make them 11 digits long. Finish the task without altering unrelated areas.
Generated: 2025-11-24 07:26:08
Status: success
Model: o3
Total Steps: 1
"""

import openpyxl
import os
import datetime


def verify_task(file_path: str) -> float:
    """Verify that every value from the 'Transaction ID' column has been copied
    to the 'Padded Transaction' column, zero-padded to 11 digits, **without**
    altering any other data in the sheet.

    Returns a progressive score between 0.0 and 1.0.
    """

    print(f"Verifying file: {file_path}\n")
    total_score = 0.0
    max_score = 1.0

    # ---------- 1.  Load workbook ----------
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        print("✓ Workbook loaded")
    except Exception as e:
        print(f"✗ Failed to load workbook: {e}")
        return 0.0  # cannot score anything if file cannot be opened

    # ---------- 2.  Locate required sheet ----------
    if "Transactions" in wb.sheetnames:
        ws = wb["Transactions"]
    else:
        ws = wb.active  # fall back to active if named sheet not present
        print(f"⚠️ 'Transactions' sheet missing – using active sheet '{ws.title}'")

    # ---------- 3.  Verify header structure ----------
    header = [c.value for c in ws[1]]
    print("Header row:", header)

    try:
        col_txn = header.index("Transaction ID")
        print(f"✓ 'Transaction ID' column found at index {col_txn}")
    except ValueError:
        print("✗ 'Transaction ID' column not found – task failed")
        return 0.0

    try:
        col_pad = header.index("Padded Transaction")
        print(f"✓ 'Padded Transaction' column found at index {col_pad}")
        total_score += 0.2  # basic structure present earns 0.2
    except ValueError:
        print("✗ 'Padded Transaction' column missing – cannot verify padding")
        col_pad = None  # continue to award 0 for padding later

    # ---------- 4.  Collect data rows ----------
    data_rows = [row for row in ws.iter_rows(min_row=2, values_only=True)
                 if not all(cell is None for cell in row)]
    num_rows = len(data_rows)
    print(f"Data rows detected: {num_rows}\n")

    # ---------- 5.  Verify padding correctness ----------
    padded_correct = 0
    if col_pad is not None and num_rows:
        for idx, row in enumerate(data_rows, start=2):  # start=2 because row 2 is first data row
            txn_val = row[col_txn]
            pad_val = row[col_pad]

            # Normalise txn value to string (remove trailing .0 for ints stored as floats)
            if isinstance(txn_val, float) and abs(txn_val - int(txn_val)) < 1e-9:
                txn_val = int(txn_val)
            txn_str = "" if txn_val is None else str(txn_val).rstrip(".0")
            expected_pad = txn_str.zfill(11) if txn_str else None

            # Normalise stored padded value similarly
            if isinstance(pad_val, float) and abs(pad_val - int(pad_val)) < 1e-9:
                pad_val = int(pad_val)
            pad_str = None if pad_val is None else str(pad_val).rstrip(".0")

            if expected_pad == pad_str:
                padded_correct += 1
            else:
                print(f"Row {idx}: expected '{expected_pad}' but found '{pad_str}'")

        padding_fraction = padded_correct / num_rows
        print(f"Padded correctness: {padded_correct}/{num_rows} ({padding_fraction:.0%})")
        total_score += 0.6 * padding_fraction  # up to 0.6 for padding work
    else:
        print("Padding verification skipped (column missing or no data rows)")

    # ---------- 6.  Verify original data unchanged (spot-check first 5 rows) ----------
    # Expected initial snapshot (from task description)
    expected_snapshot = [
        ["2023-01-03", "Alice", "12345", "250.75"],
        ["2023-01-05", "Bob", "987654321", "40"],
        ["2023-02-10", "Charlie", "1", "1000"],
        ["2023-03-15", "Dana", "55555555555", "5.55"],
        ["2023-04-22", "Eve", "67890", "300"],
    ]
    cols_to_check = ["Date", "Customer", "Transaction ID", "Amount"]
    idx_to_check = [header.index(c) for c in cols_to_check]

    cells_ok = 0
    cells_total = 0
    for r in range(min(5, num_rows)):
        row_actual = data_rows[r]
        row_expected = expected_snapshot[r]
        for j, col_idx in enumerate(idx_to_check):
            cells_total += 1
            actual_val = row_actual[col_idx]
            exp_val = row_expected[j]
            # Convert dates & floats for fair comparison
            if isinstance(actual_val, datetime.datetime):
                actual_val = actual_val.date().isoformat()
            elif isinstance(actual_val, float) and abs(actual_val - int(actual_val)) < 1e-9:
                actual_val = str(int(actual_val))
            else:
                actual_val = str(actual_val)
            if actual_val == str(exp_val):
                cells_ok += 1
            else:
                print(f"Data changed at row {r+2} col '{cols_to_check[j]}': expected '{exp_val}', got '{actual_val}'")

    if cells_total:
        unchanged_fraction = cells_ok / cells_total
        print(f"Unchanged data accuracy: {cells_ok}/{cells_total} ({unchanged_fraction:.0%})")
        total_score += 0.2 * unchanged_fraction  # up to 0.2 for not modifying other fields

    # ---------- 7.  Final score ----------
    final_score = round(min(total_score, max_score), 2)
    print(f"\nFinal score: {final_score}\n")
    return final_score


if __name__ == "__main__":
    FILE_PATH = "/home/user/copy_all_entries_from_transaction_id_to_padded_transaction_and_prepend_zeros_to_make_them_11_digits_.xlsx"
    if not os.path.exists(FILE_PATH):
        print(f"File not found: {FILE_PATH}\nREWARD: 0.0")
    else:
        reward = verify_task(FILE_PATH)
        print(f"REWARD: {reward}")
