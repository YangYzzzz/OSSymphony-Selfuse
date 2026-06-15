"""
Reward Script: Quarterly Expense Report with Department Summaries
Task ID: calc_grs_022
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Summary sheet has department SUMIF formulas
  Component 2 (0.20): Summary sheet has cross-tabulation with SUMIFS formulas
  Component 3 (0.15): Bar chart exists on Summary sheet
  Component 4 (0.15): Data validation dropdown on Department column
  Component 5 (0.15): Alternating row colors on Expenses sheet
  Component 6 (0.10): Date format DD-MMM-YYYY on Expenses dates
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_022'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Expenses sheet must exist with data
    if 'Expenses' not in wb.sheetnames:
        # Try Sheet1 as fallback
        expenses_candidates = [s for s in wb.sheetnames if s.lower() in ('expenses', 'sheet1')]
        if not expenses_candidates:
            print("CRITICAL: No Expenses/Sheet1 sheet found")
            print("REWARD: 0.0")
            return 0.0
        expenses_name = expenses_candidates[0]
    else:
        expenses_name = 'Expenses'

    ws_expenses = wb[expenses_name]

    # Precondition: Must have data rows (at least 30 transactions)
    if ws_expenses.max_row < 10:
        print(f"CRITICAL: Expenses sheet has only {ws_expenses.max_row} rows, expected 30+")
        print("REWARD: 0.0")
        return 0.0

    # Find Summary sheet
    summary_candidates = [s for s in wb.sheetnames if s.lower() in ('summary', 'sheet2')]
    if not summary_candidates:
        print("FAIL: No Summary/Sheet2 sheet found")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws_summary = wb[summary_candidates[0]]

    # Component 1: Summary sheet has department SUMIF formulas (0.25 points)
    # The golden file has SUMIF formulas in B4:B9 referencing Expenses!C:C and Expenses!F:F
    try:
        sumif_dept_count = 0
        # Scan the Summary sheet for SUMIF formulas that reference department data
        for r in range(1, ws_summary.max_row + 1):
            for c in range(1, ws_summary.max_column + 1):
                val = ws_summary.cell(row=r, column=c).value
                if isinstance(val, str) and 'SUMIF' in val.upper() and 'SUMIFS' not in val.upper():
                    sumif_dept_count += 1

        if sumif_dept_count >= 6:
            print(f"PASS: Component 1 — Found {sumif_dept_count} SUMIF formulas on Summary (0.25 pts)")
            total_score += 0.25
        elif sumif_dept_count >= 3:
            partial = 0.15
            print(f"PARTIAL: Component 1 — Found {sumif_dept_count} SUMIF formulas (partial: {partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Found only {sumif_dept_count} SUMIF formulas, expected >=6")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Cross-tabulation with SUMIFS formulas (0.20 points)
    # Golden has a dept vs category cross-tab using SUMIFS in rows 26-31
    try:
        sumifs_count = 0
        for r in range(1, ws_summary.max_row + 1):
            for c in range(1, ws_summary.max_column + 1):
                val = ws_summary.cell(row=r, column=c).value
                if isinstance(val, str) and 'SUMIFS' in val.upper():
                    sumifs_count += 1

        if sumifs_count >= 20:
            print(f"PASS: Component 2 — Found {sumifs_count} SUMIFS formulas for cross-tabulation (0.20 pts)")
            total_score += 0.20
        elif sumifs_count >= 10:
            partial = 0.10
            print(f"PARTIAL: Component 2 — Found {sumifs_count} SUMIFS formulas (partial: {partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Found only {sumifs_count} SUMIFS formulas, expected >=20 for cross-tab")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Bar chart on Summary sheet (0.15 points)
    try:
        chart_count = len(ws_summary._charts)
        if chart_count >= 1:
            # Check if it's a bar/column chart
            chart = ws_summary._charts[0]
            chart_class = chart.__class__.__name__
            if 'Bar' in chart_class:
                print(f"PASS: Component 3 — Found BarChart on Summary sheet (0.15 pts)")
                total_score += 0.15
            else:
                # Any chart type gets partial credit
                partial = 0.10
                print(f"PARTIAL: Component 3 — Found {chart_class} (not BarChart) on Summary ({partial} pts)")
                total_score += partial
        else:
            print(f"FAIL: Component 3 — No charts found on Summary sheet")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Data validation dropdown on Department column (0.15 points)
    try:
        dvs = ws_expenses.data_validations.dataValidation if ws_expenses.data_validations else []
        dept_dv_found = False
        for dv in dvs:
            if dv.type == 'list':
                # Check if it applies to column C (Department)
                sqref_str = str(dv.sqref)
                if 'C' in sqref_str:
                    dept_dv_found = True
                    print(f"PASS: Component 4 — Data validation list found on {sqref_str} (0.15 pts)")
                    total_score += 0.15
                    break

        if not dept_dv_found:
            # Check any list-type validation exists
            list_dvs = [dv for dv in dvs if dv.type == 'list']
            if list_dvs:
                partial = 0.08
                print(f"PARTIAL: Component 4 — List validation found but not on column C ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 4 — No list data validation found (found {len(dvs)} validations total)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Alternating row colors on Expenses sheet (0.15 points)
    try:
        # Check if data rows have alternating fill patterns
        # Golden: even rows (2,4,6) have FFD9E2F3, odd rows (3,5,7) have FFFFFFFF
        colored_rows = 0
        uncolored_rows = 0
        total_data_rows = min(ws_expenses.max_row, 33) - 1  # exclude header

        for r in range(2, min(ws_expenses.max_row + 1, 34)):
            cell = ws_expenses.cell(row=r, column=1)
            try:
                fill_type = cell.fill.fill_type
                if fill_type and fill_type != 'none':
                    colored_rows += 1
                else:
                    uncolored_rows += 1
            except:
                uncolored_rows += 1

        # For alternating colors, we expect roughly half colored, half not (or all colored with 2 alternating colors)
        # The key test: initial has NO colored rows, golden has colored rows
        if colored_rows >= total_data_rows * 0.4:
            # Verify alternation: check that colors alternate between at least 2 patterns
            fills = []
            for r in range(2, min(10, ws_expenses.max_row + 1)):
                cell = ws_expenses.cell(row=r, column=1)
                try:
                    fill_rgb = cell.fill.fgColor.rgb if cell.fill.fgColor else 'none'
                    fill_type = cell.fill.fill_type
                    fills.append((fill_type, fill_rgb))
                except:
                    fills.append(('none', 'none'))

            unique_fills = set(fills)
            if len(unique_fills) >= 2:
                print(f"PASS: Component 5 — Alternating row colors found ({colored_rows} colored rows, {len(unique_fills)} distinct fills) (0.15 pts)")
                total_score += 0.15
            else:
                partial = 0.08
                print(f"PARTIAL: Component 5 — Rows are colored but not alternating ({partial} pts)")
                total_score += partial
        else:
            print(f"FAIL: Component 5 — Only {colored_rows}/{total_data_rows} rows have fills (expected alternating colors)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Date format DD-MMM-YYYY (0.10 points)
    try:
        date_format_count = 0
        total_dates = 0
        for r in range(2, min(ws_expenses.max_row + 1, 34)):
            cell = ws_expenses.cell(row=r, column=1)
            if cell.value is not None:
                total_dates += 1
                fmt = cell.number_format
                # Check for DD-MMM-YYYY or similar day-month_name-year format
                if fmt and ('MMM' in fmt.upper() or 'mmm' in fmt):
                    date_format_count += 1

        if total_dates > 0 and date_format_count >= total_dates * 0.8:
            print(f"PASS: Component 6 — {date_format_count}/{total_dates} dates formatted as DD-MMM-YYYY (0.10 pts)")
            total_score += 0.10
        elif date_format_count > 0:
            partial = 0.05
            print(f"PARTIAL: Component 6 — Only {date_format_count}/{total_dates} dates formatted correctly ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 6 — No dates with DD-MMM-YYYY format (format found: {ws_expenses.cell(row=2, column=1).number_format})")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
