"""
Initial Setup: Quarterly expense report with department summaries
Task ID: calc_grs_022
Domain: libreoffice_calc

Creates a workbook with 30+ expense transactions on Sheet1 (Expenses)
and an empty Sheet2 (Summary) ready for the agent to populate.
No formulas, no charts, no alternating colors, no data validation,
no date formatting applied — those are the agent's tasks.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import date

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_022'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: Expenses ---
    ws1 = wb.active
    ws1.title = 'Expenses'

    headers = [
        'Date', 'Employee', 'Department', 'Expense Category',
        'Description', 'Amount', 'Receipts Attached', 'Approved'
    ]
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)
        ws1.cell(row=1, column=col).font = Font(bold=True)

    # Departments and categories for realistic data
    departments = ['Engineering', 'Marketing', 'Sales', 'Finance', 'HR', 'Operations']
    categories = ['Travel', 'Meals & Entertainment', 'Office Supplies', 'Software', 'Training', 'Equipment']

    # 32 realistic expense transactions
    transactions = [
        (date(2025, 1, 8), 'Sarah Chen', 'Engineering', 'Software', 'JetBrains IDE annual license', 249.00, 'Y', 'Y'),
        (date(2025, 1, 12), 'Marcus Johnson', 'Marketing', 'Meals & Entertainment', 'Client dinner at Nobu downtown', 387.50, 'Y', 'Y'),
        (date(2025, 1, 15), 'Priya Patel', 'Sales', 'Travel', 'Flight to Chicago for trade show', 542.00, 'Y', 'Y'),
        (date(2025, 1, 18), 'David Kim', 'Finance', 'Office Supplies', 'Printer toner cartridges x3', 189.75, 'Y', 'Y'),
        (date(2025, 1, 22), 'Elena Rodriguez', 'HR', 'Training', 'Diversity & inclusion workshop', 1200.00, 'Y', 'Y'),
        (date(2025, 1, 25), 'James Wright', 'Operations', 'Equipment', 'Warehouse barcode scanners x2', 678.00, 'Y', 'Y'),
        (date(2025, 2, 3), 'Sarah Chen', 'Engineering', 'Travel', 'Uber rides to AWS Summit', 87.40, 'Y', 'Y'),
        (date(2025, 2, 5), 'Marcus Johnson', 'Marketing', 'Meals & Entertainment', 'Team lunch for campaign launch', 215.00, 'N', 'Y'),
        (date(2025, 2, 8), 'Lisa Thompson', 'Marketing', 'Travel', 'Hotel for content creator meetup', 329.00, 'Y', 'Y'),
        (date(2025, 2, 12), 'Priya Patel', 'Sales', 'Meals & Entertainment', 'Prospect dinner at Eleven Madison', 456.30, 'Y', 'Y'),
        (date(2025, 2, 14), 'David Kim', 'Finance', 'Software', 'QuickBooks subscription renewal', 150.00, 'Y', 'Y'),
        (date(2025, 2, 18), 'Angela Foster', 'HR', 'Office Supplies', 'Ergonomic keyboard and mouse sets', 324.80, 'Y', 'N'),
        (date(2025, 2, 21), 'James Wright', 'Operations', 'Travel', 'Mileage reimbursement Feb', 198.50, 'Y', 'Y'),
        (date(2025, 2, 24), 'Tom Richards', 'Engineering', 'Equipment', 'External monitors x2 for new hires', 890.00, 'Y', 'Y'),
        (date(2025, 2, 27), 'Marcus Johnson', 'Marketing', 'Meals & Entertainment', 'VIP client cocktail event', 1450.00, 'Y', 'Y'),
        (date(2025, 3, 3), 'Sarah Chen', 'Engineering', 'Training', 'Kubernetes certification course', 395.00, 'Y', 'Y'),
        (date(2025, 3, 5), 'Priya Patel', 'Sales', 'Office Supplies', 'Business cards reprint 500ct', 85.00, 'Y', 'Y'),
        (date(2025, 3, 7), 'Lisa Thompson', 'Marketing', 'Software', 'Adobe Creative Cloud annual', 599.88, 'Y', 'Y'),
        (date(2025, 3, 10), 'David Kim', 'Finance', 'Training', 'CPA continuing education seminar', 275.00, 'Y', 'Y'),
        (date(2025, 3, 12), 'Elena Rodriguez', 'HR', 'Travel', 'Flight to LA for recruiting fair', 412.00, 'Y', 'Y'),
        (date(2025, 3, 14), 'James Wright', 'Operations', 'Equipment', 'Safety vests and hard hats bulk', 520.00, 'Y', 'Y'),
        (date(2025, 3, 17), 'Tom Richards', 'Engineering', 'Software', 'GitHub Enterprise seats x5', 210.00, 'Y', 'Y'),
        (date(2025, 3, 19), 'Marcus Johnson', 'Marketing', 'Meals & Entertainment', 'Influencer brunch event catering', 890.00, 'Y', 'N'),
        (date(2025, 3, 20), 'Angela Foster', 'HR', 'Office Supplies', 'Standing desk converter x4', 1196.00, 'Y', 'Y'),
        (date(2025, 3, 21), 'Priya Patel', 'Sales', 'Travel', 'Rental car for client visits', 275.00, 'Y', 'Y'),
        (date(2025, 3, 22), 'David Kim', 'Finance', 'Meals & Entertainment', 'Quarterly team lunch', 168.50, 'Y', 'Y'),
        (date(2025, 3, 24), 'Sarah Chen', 'Engineering', 'Office Supplies', 'Mechanical keyboards x3', 447.00, 'N', 'Y'),
        (date(2025, 3, 25), 'Lisa Thompson', 'Marketing', 'Travel', 'Taxi to press conference', 62.00, 'Y', 'Y'),
        (date(2025, 3, 26), 'James Wright', 'Operations', 'Training', 'OSHA compliance certification', 350.00, 'Y', 'Y'),
        (date(2025, 3, 27), 'Elena Rodriguez', 'HR', 'Meals & Entertainment', 'New hire welcome lunch x6', 234.00, 'Y', 'Y'),
        (date(2025, 3, 28), 'Tom Richards', 'Engineering', 'Travel', 'Conference registration PyCon', 450.00, 'Y', 'Y'),
        (date(2025, 3, 31), 'Marcus Johnson', 'Marketing', 'Equipment', 'Ring light and tripod for content', 189.99, 'Y', 'Y'),
    ]

    for r, row_data in enumerate(transactions, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws1.column_dimensions['A'].width = 14
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 16
    ws1.column_dimensions['D'].width = 22
    ws1.column_dimensions['E'].width = 40
    ws1.column_dimensions['F'].width = 14
    ws1.column_dimensions['G'].width = 20
    ws1.column_dimensions['H'].width = 12

    # Format Amount column as currency
    for r in range(2, len(transactions) + 2):
        ws1.cell(row=r, column=6).number_format = '$#,##0.00'

    # --- Sheet2: Summary (empty — agent must build this) ---
    ws2 = wb.create_sheet('Summary')
    # Leave empty for agent to populate with SUMIF formulas, cross-tab, and chart

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
