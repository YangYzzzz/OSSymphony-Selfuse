"""
Initial Setup: Conference room booking calendar - raw data before organization
Task ID: calc_grs_027
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_027'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Raw Bookings ---
    ws1 = wb.active
    ws1.title = 'Raw Bookings'

    headers = ['Date', 'Start Time', 'End Time', 'Room', 'Organizer',
               'Department', 'Meeting Name', 'Number of Attendees']
    header_font = Font(bold=True, size=11)
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font

    # Realistic booking data for 4 weeks (March 2026)
    bookings = [
        # Week 1
        ['2026-03-02', '9:00 AM', '10:00 AM', 'Room A', 'Sarah Chen', 'Engineering', 'Sprint Planning', 12],
        ['2026-03-02', '10:00 AM', '11:00 AM', 'Room B', 'David Park', 'Sales', 'Q1 Pipeline Review', 8],
        ['2026-03-02', '1:00 PM', '3:00 PM', 'Room A', 'Rachel Torres', 'HR', 'Benefits Enrollment Session', 25],
        ['2026-03-02', '2:00 PM', '3:00 PM', 'Room B', 'Michael Foster', 'Management', 'Budget Review', 5],
        ['2026-03-03', '8:00 AM', '9:00 AM', 'Room A', 'Lisa Wang', 'Engineering', 'Code Review Sync', 6],
        ['2026-03-03', '11:00 AM', '12:00 PM', 'Room B', 'James O\'Brien', 'External', 'Vendor Demo - CloudStack', 10],
        ['2026-03-03', '3:00 PM', '5:00 PM', 'Room A', 'Sarah Chen', 'Engineering', 'Architecture Workshop', 15],
        ['2026-03-04', '9:00 AM', '10:00 AM', 'Room B', 'Amanda Lewis', 'Sales', 'Client Pitch - Acme Corp', 4],
        ['2026-03-04', '10:00 AM', '12:00 PM', 'Room A', 'Rachel Torres', 'HR', 'New Hire Orientation', 8],
        ['2026-03-04', '1:00 PM', '2:00 PM', 'Room B', 'Michael Foster', 'Management', 'Leadership Sync', 6],
        ['2026-03-05', '9:00 AM', '11:00 AM', 'Room A', 'David Park', 'Sales', 'Sales Training Workshop', 14],
        ['2026-03-05', '2:00 PM', '3:00 PM', 'Room B', 'Lisa Wang', 'Engineering', 'Tech Debt Discussion', 7],
        ['2026-03-06', '10:00 AM', '11:00 AM', 'Room A', 'James O\'Brien', 'External', 'Partner Meeting - Nexus', 5],
        ['2026-03-06', '1:00 PM', '2:00 PM', 'Room B', 'Amanda Lewis', 'Sales', 'Forecast Update', 3],
        # Week 2
        ['2026-03-09', '9:00 AM', '10:00 AM', 'Room A', 'Sarah Chen', 'Engineering', 'Sprint Retrospective', 12],
        ['2026-03-09', '11:00 AM', '12:00 PM', 'Room B', 'Rachel Torres', 'HR', 'Policy Review Meeting', 4],
        ['2026-03-09', '2:00 PM', '4:00 PM', 'Room A', 'Michael Foster', 'Management', 'Quarterly Strategy', 8],
        ['2026-03-10', '8:00 AM', '9:00 AM', 'Room B', 'Lisa Wang', 'Engineering', 'Morning Standup Extended', 10],
        ['2026-03-10', '10:00 AM', '11:00 AM', 'Room A', 'David Park', 'Sales', 'Territory Planning', 6],
        ['2026-03-10', '1:00 PM', '2:00 PM', 'Room B', 'James O\'Brien', 'External', 'Audit Preparation', 5],
        ['2026-03-11', '9:00 AM', '11:00 AM', 'Room A', 'Rachel Torres', 'HR', 'Interview Panel - Sr Dev', 4],
        ['2026-03-11', '2:00 PM', '3:00 PM', 'Room B', 'Amanda Lewis', 'Sales', 'Demo Prep Session', 3],
        ['2026-03-12', '10:00 AM', '12:00 PM', 'Room A', 'Sarah Chen', 'Engineering', 'Hackathon Kickoff', 20],
        ['2026-03-12', '1:00 PM', '2:00 PM', 'Room B', 'Michael Foster', 'Management', 'Performance Reviews', 2],
        ['2026-03-13', '9:00 AM', '10:00 AM', 'Room A', 'Lisa Wang', 'Engineering', 'Release Planning', 8],
        ['2026-03-13', '3:00 PM', '4:00 PM', 'Room B', 'David Park', 'Sales', 'Win/Loss Analysis', 5],
        # Week 3
        ['2026-03-16', '9:00 AM', '10:00 AM', 'Room A', 'Sarah Chen', 'Engineering', 'Sprint Planning', 12],
        ['2026-03-16', '10:00 AM', '11:00 AM', 'Room B', 'Rachel Torres', 'HR', 'Compensation Review', 3],
        ['2026-03-16', '2:00 PM', '3:00 PM', 'Room A', 'Amanda Lewis', 'Sales', 'Pipeline Deep Dive', 7],
        ['2026-03-17', '8:00 AM', '10:00 AM', 'Room B', 'James O\'Brien', 'External', 'Client Workshop - TechVentures', 15],
        ['2026-03-17', '11:00 AM', '12:00 PM', 'Room A', 'Michael Foster', 'Management', 'Board Prep Session', 4],
        ['2026-03-17', '1:00 PM', '3:00 PM', 'Room B', 'Lisa Wang', 'Engineering', 'System Design Review', 9],
        ['2026-03-18', '9:00 AM', '10:00 AM', 'Room A', 'David Park', 'Sales', 'Regional Sync - West', 5],
        ['2026-03-18', '2:00 PM', '4:00 PM', 'Room B', 'Rachel Torres', 'HR', 'Training: DEI Workshop', 30],
        ['2026-03-19', '10:00 AM', '11:00 AM', 'Room A', 'Sarah Chen', 'Engineering', 'Incident Post-Mortem', 8],
        ['2026-03-19', '1:00 PM', '2:00 PM', 'Room B', 'Amanda Lewis', 'Sales', 'Proposal Review', 4],
        ['2026-03-20', '9:00 AM', '11:00 AM', 'Room A', 'Michael Foster', 'Management', 'All-Hands Prep', 6],
        ['2026-03-20', '2:00 PM', '3:00 PM', 'Room B', 'Lisa Wang', 'Engineering', 'API Design Session', 5],
        # Week 4
        ['2026-03-23', '9:00 AM', '10:00 AM', 'Room A', 'Sarah Chen', 'Engineering', 'Sprint Retrospective', 12],
        ['2026-03-23', '11:00 AM', '12:00 PM', 'Room B', 'David Park', 'Sales', 'Monthly Forecast', 8],
        ['2026-03-23', '2:00 PM', '3:00 PM', 'Room A', 'Rachel Torres', 'HR', 'Exit Interview Debrief', 3],
        ['2026-03-24', '8:00 AM', '9:00 AM', 'Room B', 'Lisa Wang', 'Engineering', 'Deployment Checklist', 6],
        ['2026-03-24', '10:00 AM', '12:00 PM', 'Room A', 'James O\'Brien', 'External', 'Investor Presentation', 10],
        ['2026-03-24', '1:00 PM', '2:00 PM', 'Room B', 'Michael Foster', 'Management', 'Risk Assessment', 4],
        ['2026-03-25', '9:00 AM', '10:00 AM', 'Room A', 'Amanda Lewis', 'Sales', 'Competitive Analysis', 5],
        ['2026-03-25', '2:00 PM', '4:00 PM', 'Room B', 'Sarah Chen', 'Engineering', 'Load Testing Review', 7],
        ['2026-03-26', '10:00 AM', '11:00 AM', 'Room A', 'Rachel Torres', 'HR', 'Team Building Planning', 4],
        ['2026-03-26', '1:00 PM', '3:00 PM', 'Room B', 'David Park', 'Sales', 'Account Strategy Session', 6],
        ['2026-03-27', '9:00 AM', '10:00 AM', 'Room A', 'Michael Foster', 'Management', 'Month-End Review', 5],
        ['2026-03-27', '2:00 PM', '3:00 PM', 'Room B', 'Lisa Wang', 'Engineering', 'Tooling Evaluation', 4],
    ]

    for r, row_data in enumerate(bookings, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws1.column_dimensions['A'].width = 14
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 12
    ws1.column_dimensions['D'].width = 10
    ws1.column_dimensions['E'].width = 18
    ws1.column_dimensions['F'].width = 14
    ws1.column_dimensions['G'].width = 32
    ws1.column_dimensions['H'].width = 20

    # --- Sheet 2: Rooms ---
    ws2 = wb.create_sheet('Rooms')
    room_headers = ['Room Name', 'Capacity', 'Equipment', 'Floor', 'Notes']
    for col, h in enumerate(room_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font

    room_data = [
        ['Room A', 30, 'Projector, Whiteboard, Video Conf', '2nd Floor', 'Large conference room - ideal for all-hands'],
        ['Room B', 12, 'TV Screen, Whiteboard', '2nd Floor', 'Medium room - good for team meetings'],
    ]
    for r, row_data in enumerate(room_data, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 35
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 45

    # --- Sheet 3: Department Colors (reference) ---
    ws3 = wb.create_sheet('Department Colors')
    color_headers = ['Department', 'Color']
    for col, h in enumerate(color_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font

    dept_colors = [
        ['Sales', 'Blue'],
        ['HR', 'Green'],
        ['Engineering', 'Orange'],
        ['Management', 'Red'],
        ['External', 'Purple'],
    ]
    for r, row_data in enumerate(dept_colors, 2):
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r, column=c, value=val)

    ws3.column_dimensions['A'].width = 15
    ws3.column_dimensions['B'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
