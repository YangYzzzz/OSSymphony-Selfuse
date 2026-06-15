"""
Reward Script: Conference Room Booking Calendar
Task ID: calc_grs_027
Domain: libreoffice_calc
Scoring:
  Component 1: Weekly calendar sheets exist with proper structure (0.20)
  Component 2: Color-coded bookings using department colors (0.20)
  Component 3: Merged cells for multi-hour bookings (0.15)
  Component 4: Booking Log sheet with columns and data validation (0.15)
  Component 5: Summary sheet with department hours totals (0.15)
  Component 6: Pie chart on Summary sheet (0.15)
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_027'

# Department color map (ARGB)
DEPT_COLORS = {
    'FF4472C4': 'Sales',      # Blue
    'FF70AD47': 'HR',          # Green
    'FFED7D31': 'Engineering', # Orange
    'FFFF0000': 'Management',  # Red
    'FF7030A0': 'External',    # Purple
}

EXPECTED_DEPTS = {'Sales', 'HR', 'Engineering', 'Management', 'External'}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    sheet_names = wb.sheetnames

    # ================================================================
    # Component 1: Weekly calendar sheets exist with proper structure (0.20)
    # Initial has no Week sheets; Golden has Week 1-4 with time rows
    # ================================================================
    try:
        week_sheets_found = 0
        valid_week_sheets = 0
        for week_num in range(1, 5):
            week_name = f'Week {week_num}'
            if week_name in sheet_names:
                week_sheets_found += 1
                ws = wb[week_name]
                # Check: has time slots in column A (8:00 AM through some afternoon hours)
                time_slots = []
                for row_idx in range(1, ws.max_row + 1):
                    val = ws.cell(row=row_idx, column=1).value
                    if val and 'AM' in str(val) or (val and 'PM' in str(val)):
                        time_slots.append(str(val))
                # Need at least 8 time slots (8am-3pm minimum)
                if len(time_slots) >= 8:
                    valid_week_sheets += 1

        if valid_week_sheets >= 4:
            print(f"PASS: Component 1 — All 4 week sheets found with time slots (0.20 pts)")
            total_score += 0.20
        elif valid_week_sheets >= 2:
            partial = 0.10
            print(f"PARTIAL: Component 1 — {valid_week_sheets}/4 valid week sheets ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {week_sheets_found} week sheets found, {valid_week_sheets} valid")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ================================================================
    # Component 2: Color-coded bookings using department colors (0.20)
    # Initial has no colored booking cells on calendar; Golden does
    # ================================================================
    try:
        total_colored_bookings = 0
        dept_colors_used = set()

        for week_num in range(1, 5):
            week_name = f'Week {week_num}'
            if week_name not in sheet_names:
                continue
            ws = wb[week_name]
            # Scan booking area (rows 4+, cols B-K = 2-11)
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=2, max_col=11):
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue
                    try:
                        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb in DEPT_COLORS:
                            total_colored_bookings += 1
                            dept_colors_used.add(DEPT_COLORS[cell.fill.fgColor.rgb])
                    except:
                        pass

        if total_colored_bookings >= 20 and len(dept_colors_used) >= 4:
            print(f"PASS: Component 2 — {total_colored_bookings} colored bookings across {len(dept_colors_used)} departments (0.20 pts)")
            total_score += 0.20
        elif total_colored_bookings >= 10 and len(dept_colors_used) >= 3:
            partial = 0.12
            print(f"PARTIAL: Component 2 — {total_colored_bookings} colored bookings, {len(dept_colors_used)} depts ({partial} pts)")
            total_score += partial
        elif total_colored_bookings >= 5:
            partial = 0.06
            print(f"PARTIAL: Component 2 — {total_colored_bookings} colored bookings ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {total_colored_bookings} colored bookings found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ================================================================
    # Component 3: Merged cells for multi-hour bookings (0.15)
    # Initial has no merged cells; Golden has merged cells on calendar
    # ================================================================
    try:
        total_merges_in_calendar = 0
        for week_num in range(1, 5):
            week_name = f'Week {week_num}'
            if week_name not in sheet_names:
                continue
            ws = wb[week_name]
            for merged_range in ws.merged_cells.ranges:
                mr_str = str(merged_range)
                # Check if merge is in the booking area (not header merges in row 1-2)
                # Booking area merges span at least 2 rows in cols B-K, starting row 4+
                bounds = merged_range.bounds  # (min_col, min_row, max_col, max_row)
                min_col, min_row, max_col, max_row = bounds
                if min_row >= 4 and min_col >= 2 and max_col <= 11:
                    # This is a booking-area merge (multi-hour meeting)
                    total_merges_in_calendar += 1

        if total_merges_in_calendar >= 5:
            print(f"PASS: Component 3 — {total_merges_in_calendar} merged booking ranges found (0.15 pts)")
            total_score += 0.15
        elif total_merges_in_calendar >= 2:
            partial = 0.08
            print(f"PARTIAL: Component 3 — {total_merges_in_calendar} merged ranges ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {total_merges_in_calendar} merged booking ranges found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ================================================================
    # Component 4: Booking Log sheet with proper columns and data validation (0.15)
    # Initial has no Booking Log sheet; Golden has it with validation
    # ================================================================
    try:
        booking_log_score = 0.0

        if 'Booking Log' in sheet_names:
            ws_log = wb['Booking Log']
            # Check required columns in header row
            headers = []
            for col in range(1, ws_log.max_column + 1):
                val = ws_log.cell(row=1, column=col).value
                if val:
                    headers.append(str(val).strip())

            required_headers = {'Date', 'Room', 'Organizer', 'Department', 'Meeting Name'}
            found_headers = set(headers)
            matched = required_headers.intersection(found_headers)

            if len(matched) >= 5:
                booking_log_score += 0.07
                print(f"  Booking Log headers OK: {matched}")
            else:
                print(f"  Booking Log headers incomplete: found {found_headers}, need {required_headers}")

            # Check data validation on Department column
            has_validation = False
            if ws_log.data_validations:
                for dv in ws_log.data_validations.dataValidation:
                    if dv.type == 'list':
                        has_validation = True
                        break

            if has_validation:
                booking_log_score += 0.04
                print(f"  Booking Log data validation present")
            else:
                print(f"  Booking Log data validation missing")

            # Check that booking log has data rows
            if ws_log.max_row >= 10:
                booking_log_score += 0.04
                print(f"  Booking Log has {ws_log.max_row - 1} data rows")
            else:
                print(f"  Booking Log has only {ws_log.max_row - 1} data rows")

        if booking_log_score > 0:
            print(f"PASS: Component 4 — Booking Log sheet ({booking_log_score:.2f} pts)")
            total_score += booking_log_score
        else:
            print(f"FAIL: Component 4 — No valid Booking Log sheet found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ================================================================
    # Component 5: Summary sheet with department hours totals (0.15)
    # Initial has no Summary sheet; Golden has dept hour tallies
    # ================================================================
    try:
        summary_score = 0.0

        if 'Summary' in sheet_names:
            ws_sum = wb['Summary']

            # Check department names present in Summary
            depts_found = set()
            hours_found = {}
            for row in ws_sum.iter_rows(min_row=2, max_row=ws_sum.max_row, min_col=1, max_col=2):
                dept_val = row[0].value
                hours_val = row[1].value
                if dept_val and str(dept_val).strip() in EXPECTED_DEPTS:
                    depts_found.add(str(dept_val).strip())
                    if hours_val is not None and isinstance(hours_val, (int, float)) and hours_val > 0:
                        hours_found[str(dept_val).strip()] = hours_val

            if len(depts_found) >= 4 and len(hours_found) >= 4:
                summary_score += 0.10
                print(f"  Summary has {len(depts_found)} departments with hours: {hours_found}")
            elif len(depts_found) >= 2:
                summary_score += 0.05
                print(f"  Summary has {len(depts_found)} departments (partial)")
            else:
                print(f"  Summary missing departments: found {depts_found}")

            # Check total row exists
            has_total = False
            for row in ws_sum.iter_rows(min_row=2, max_row=ws_sum.max_row, min_col=1, max_col=2):
                val = row[0].value
                if val and 'TOTAL' in str(val).upper():
                    has_total = True
                    break
            if has_total:
                summary_score += 0.05
                print(f"  Summary has TOTAL row")

        if summary_score > 0:
            print(f"PASS: Component 5 — Summary sheet ({summary_score:.2f} pts)")
            total_score += summary_score
        else:
            print(f"FAIL: Component 5 — No valid Summary sheet found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # ================================================================
    # Component 6: Pie chart on Summary sheet (0.15)
    # Initial has no charts; Golden has a pie chart
    # ================================================================
    try:
        has_pie_chart = False
        if 'Summary' in sheet_names:
            ws_sum = wb['Summary']
            charts = ws_sum._charts
            if len(charts) >= 1:
                for chart in charts:
                    chart_class = chart.__class__.__name__
                    if 'Pie' in chart_class:
                        has_pie_chart = True
                        break

        if has_pie_chart:
            print(f"PASS: Component 6 — Pie chart found on Summary sheet (0.15 pts)")
            total_score += 0.15
        else:
            # Check any chart exists on Summary
            if 'Summary' in sheet_names and len(wb['Summary']._charts) >= 1:
                print(f"PARTIAL: Component 6 — Chart found but not a pie chart (0.07 pts)")
                total_score += 0.07
            else:
                print(f"FAIL: Component 6 — No chart found on Summary sheet")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
