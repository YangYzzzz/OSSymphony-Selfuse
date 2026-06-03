"""
Reward Script: Data validation dropdown for deal stages in pipeline tracker
Task ID: calc_sales_046
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): A data validation exists on the Pipeline sheet
  Component 2 (0.3): Validation is type 'list' referencing Config!$A$2:$A$7
  Component 3 (0.4): Validation applies to all cells C2:C6
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_046'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice document."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Pipeline sheet must exist
    if 'Pipeline' not in wb.sheetnames:
        print("FAIL: 'Pipeline' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Pipeline']

    # Get data validations on Pipeline sheet
    validations = []
    if ws.data_validations and ws.data_validations.dataValidation:
        validations = list(ws.data_validations.dataValidation)

    # Component 1: At least one data validation exists on Pipeline sheet (0.3 points)
    try:
        if len(validations) > 0:
            print(f"PASS: Component 1 — {len(validations)} data validation(s) found on Pipeline sheet (0.3 pts)")
            total_score += 0.3
        else:
            print("FAIL: Component 1 — No data validations found on Pipeline sheet")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Validation is type 'list' and references Config!$A$2:$A$7 (0.3 points)
    # Find the relevant validation (type=list with a formula referencing Config sheet)
    list_dv = None
    try:
        for dv in validations:
            if dv.type == 'list':
                list_dv = dv
                break

        if list_dv is not None:
            formula = str(list_dv.formula1).strip() if list_dv.formula1 else ''
            # Normalize: remove leading '=' if present, compare case-insensitively
            norm_formula = formula.lstrip('=').replace(' ', '').upper()
            expected_variants = [
                "CONFIG!$A$2:$A$7",
                "CONFIG!A2:A7",
                "'CONFIG'!$A$2:$A$7",
                "'CONFIG'!A2:A7",
            ]
            formula_ok = any(norm_formula == v.upper() for v in expected_variants)

            if formula_ok:
                print(f"PASS: Component 2 — Validation type=list, formula1={formula} matches Config!$A$2:$A$7 (0.3 pts)")
                total_score += 0.3
            else:
                # Also accept inline list of the six stages as a valid alternative
                stages = ['Lead', 'Qualified', 'Proposal', 'Negotiation', 'Closed Won', 'Closed Lost']
                stages_str = ','.join(stages)
                # formula1 for inline list is like '"Lead,Qualified,..."'
                inline_formula = norm_formula.strip('"').strip("'")
                inline_expected = stages_str.upper()
                if inline_formula == inline_expected:
                    print(f"PASS: Component 2 — Validation type=list with inline stage values (0.3 pts)")
                    total_score += 0.3
                else:
                    print(f"FAIL: Component 2 — Validation formula1={formula}, expected reference to Config!$A$2:$A$7")
        else:
            print("FAIL: Component 2 — No list-type data validation found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Validation covers cells C2 through C6 (0.4 points)
    try:
        if list_dv is not None:
            sqref_str = str(list_dv.sqref).upper().replace(' ', '')
            # The sqref could be "C2:C6" or individual cells "C2 C3 C4 C5 C6"
            # Check that all required cells are covered
            required_cells = {'C2', 'C3', 'C4', 'C5', 'C6'}
            covered_cells = set()

            # Parse sqref ranges
            for part in sqref_str.split():
                if ':' in part:
                    # Range like C2:C6
                    from openpyxl.utils import range_boundaries
                    min_col, min_row, max_col, max_row = range_boundaries(part)
                    for r in range(min_row, max_row + 1):
                        for c_idx in range(min_col, max_col + 1):
                            from openpyxl.utils import get_column_letter
                            cell_ref = f"{get_column_letter(c_idx)}{r}"
                            covered_cells.add(cell_ref.upper())
                else:
                    covered_cells.add(part.upper())

            if required_cells.issubset(covered_cells):
                print(f"PASS: Component 3 — Validation covers all required cells C2:C6 (sqref={list_dv.sqref}) (0.4 pts)")
                total_score += 0.4
            else:
                missing = required_cells - covered_cells
                print(f"FAIL: Component 3 — Validation sqref={list_dv.sqref}, missing cells: {missing}")
        else:
            print("FAIL: Component 3 — No list-type data validation found to check range")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
