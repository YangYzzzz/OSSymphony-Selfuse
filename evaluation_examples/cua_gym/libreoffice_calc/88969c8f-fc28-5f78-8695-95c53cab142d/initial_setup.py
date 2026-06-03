"""
Initial Setup: Create a sales pipeline tracker with deal data and a Config sheet with stage names.
Task ID: calc_sales_046
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_046'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Pipeline ---
    ws1 = wb.active
    ws1.title = 'Pipeline'

    # Headers
    headers = ['Deal', 'Account', 'Stage', 'Value']
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows - Deal ID, Account, Stage (empty), Value
    pipeline_data = [
        ['D001', 'Acme Corp', None, 125000],
        ['D002', 'Beta Inc', None, 87500],
        ['D003', 'Gamma LLC', None, 215000],
        ['D004', 'Delta Co', None, 64000],
        ['D005', 'Epsilon Ltd', None, 178500],
    ]
    for r, row_data in enumerate(pipeline_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)

    # Format Value column as currency
    for r in range(2, 7):
        ws1.cell(row=r, column=4).number_format = '$#,##0'

    # Set column widths for readability
    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 16
    ws1.column_dimensions['D'].width = 14

    # --- Sheet 2: Config ---
    ws2 = wb.create_sheet('Config')
    ws2.cell(row=1, column=1, value='Stages')
    ws2['A1'].font = Font(bold=True)

    stages = ['Lead', 'Qualified', 'Proposal', 'Negotiation', 'Closed Won', 'Closed Lost']
    for i, stage in enumerate(stages, 2):
        ws2.cell(row=i, column=1, value=stage)

    ws2.column_dimensions['A'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
