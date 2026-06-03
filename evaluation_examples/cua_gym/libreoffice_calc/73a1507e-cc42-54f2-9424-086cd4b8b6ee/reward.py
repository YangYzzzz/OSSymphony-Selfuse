"""
Reward Script: Reorder sheets alphabetically
Task ID: calc_ps_073
Domain: libreoffice_calc
Scoring:
  Component 1 (0.6) — Sheets in correct alphabetical order
  Component 2 (0.15) — All 4 expected sheets present
  Component 3 (0.25) — Data preserved in each sheet (spot-check key cells)
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_073'

EXPECTED_ORDER = ['Apple', 'Banana', 'Mango', 'Zebra']


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    actual_names = wb.sheetnames
    print(f"INFO: Sheet names found: {actual_names}")

    # Component 1: Sheets in correct alphabetical order (0.6 points)
    # This is the core task — reorder sheets to alphabetical.
    # Initial order is ['Zebra', 'Apple', 'Mango', 'Banana'] which does NOT match.
    # Golden order is ['Apple', 'Banana', 'Mango', 'Zebra'] which DOES match.
    try:
        if actual_names == EXPECTED_ORDER:
            print(f"PASS: Component 1 — Sheets in correct order {actual_names} (0.6 pts)")
            total_score += 0.6
        else:
            # Partial credit: count how many sheets are in the correct position
            correct_positions = sum(1 for a, e in zip(actual_names, EXPECTED_ORDER) if a == e)
            if correct_positions >= 2 and len(actual_names) == 4:
                partial = 0.6 * (correct_positions / 4)
                print(f"PARTIAL: Component 1 — {correct_positions}/4 sheets in correct position ({partial:.2f} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 1 — Expected {EXPECTED_ORDER}, found {actual_names}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All 4 expected sheets present with correct names (0.15 points)
    # This checks that no sheets were lost or renamed during reordering.
    # The initial file also has all 4 sheets, BUT only award points if Component 1
    # scored > 0 (i.e., at least some reordering happened). This ensures initial_env
    # scores 0 because Component 1 fails completely (0 correct positions in initial order).
    try:
        present = set(actual_names)
        expected_set = set(EXPECTED_ORDER)
        # Only award if sheets are in different order than initial (i.e., reordering happened)
        initial_order = ['Zebra', 'Apple', 'Mango', 'Banana']
        if present == expected_set and len(actual_names) == 4 and actual_names != initial_order:
            print(f"PASS: Component 2 — All 4 sheets present and reordered (0.15 pts)")
            total_score += 0.15
        elif present == expected_set and actual_names == initial_order:
            print(f"FAIL: Component 2 — All sheets present but still in initial order (no reordering)")
        else:
            missing = expected_set - present
            extra = present - expected_set
            print(f"FAIL: Component 2 — Missing: {missing}, Extra: {extra}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Data preservation (0.25 points)
    # Verify key cells in each sheet still have correct data after reordering.
    # Only award if order has changed (anchored to task change).
    # We check specific cell values that exist in golden but won't score if order is initial.
    try:
        if actual_names == initial_order:
            # No reordering happened — don't award data preservation points
            print(f"FAIL: Component 3 — No reordering detected, skipping data check")
        else:
            data_checks_passed = 0
            data_checks_total = 4

            # Check Apple sheet: A2 should be 'Fuji'
            if 'Apple' in wb.sheetnames:
                ws = wb['Apple']
                if ws.cell(row=2, column=1).value == 'Fuji':
                    data_checks_passed += 1
                    print(f"  PASS: Apple!A2 = 'Fuji'")
                else:
                    print(f"  FAIL: Apple!A2 = {ws.cell(row=2, column=1).value}, expected 'Fuji'")
            else:
                print(f"  FAIL: Apple sheet not found")

            # Check Banana sheet: A2 should be 'Cavendish'
            if 'Banana' in wb.sheetnames:
                ws = wb['Banana']
                if ws.cell(row=2, column=1).value == 'Cavendish':
                    data_checks_passed += 1
                    print(f"  PASS: Banana!A2 = 'Cavendish'")
                else:
                    print(f"  FAIL: Banana!A2 = {ws.cell(row=2, column=1).value}, expected 'Cavendish'")
            else:
                print(f"  FAIL: Banana sheet not found")

            # Check Mango sheet: A2 should be 'Alphonso'
            if 'Mango' in wb.sheetnames:
                ws = wb['Mango']
                if ws.cell(row=2, column=1).value == 'Alphonso':
                    data_checks_passed += 1
                    print(f"  PASS: Mango!A2 = 'Alphonso'")
                else:
                    print(f"  FAIL: Mango!A2 = {ws.cell(row=2, column=1).value}, expected 'Alphonso'")
            else:
                print(f"  FAIL: Mango sheet not found")

            # Check Zebra sheet: A2 should be 'Z-001'
            if 'Zebra' in wb.sheetnames:
                ws = wb['Zebra']
                if ws.cell(row=2, column=1).value == 'Z-001':
                    data_checks_passed += 1
                    print(f"  PASS: Zebra!A2 = 'Z-001'")
                else:
                    print(f"  FAIL: Zebra!A2 = {ws.cell(row=2, column=1).value}, expected 'Z-001'")
            else:
                print(f"  FAIL: Zebra sheet not found")

            if data_checks_passed == data_checks_total:
                print(f"PASS: Component 3 — All data preserved ({data_checks_passed}/{data_checks_total}) (0.25 pts)")
                total_score += 0.25
            elif data_checks_passed > 0:
                partial = 0.25 * (data_checks_passed / data_checks_total)
                print(f"PARTIAL: Component 3 — {data_checks_passed}/{data_checks_total} data checks passed ({partial:.2f} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 3 — No data checks passed")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
