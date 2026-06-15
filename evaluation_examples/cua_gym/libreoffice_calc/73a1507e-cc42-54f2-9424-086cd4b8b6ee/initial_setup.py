"""
Initial Setup: Reorder sheets alphabetically
Task ID: calc_ps_073
Domain: libreoffice_calc
Sheets start in order: Zebra, Apple, Mango, Banana (NOT alphabetical)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_073'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Zebra ---
    ws1 = wb.active
    ws1.title = 'Zebra'
    headers = ['Animal ID', 'Species', 'Habitat', 'Weight (kg)', 'Status']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    data = [
        ['Z-001', 'Plains Zebra', 'Savanna', 350, 'Healthy'],
        ['Z-002', 'Grevy\'s Zebra', 'Grassland', 410, 'Under observation'],
        ['Z-003', 'Mountain Zebra', 'Mountain slopes', 280, 'Healthy'],
        ['Z-004', 'Plains Zebra', 'Savanna', 320, 'Recovering'],
        ['Z-005', 'Grevy\'s Zebra', 'Semi-arid', 390, 'Healthy'],
        ['Z-006', 'Plains Zebra', 'Woodland', 345, 'Healthy'],
        ['Z-007', 'Mountain Zebra', 'Rocky terrain', 275, 'Under observation'],
        ['Z-008', 'Plains Zebra', 'Savanna', 360, 'Healthy'],
        ['Z-009', 'Grevy\'s Zebra', 'Grassland', 425, 'Healthy'],
        ['Z-010', 'Mountain Zebra', 'Highland plateau', 290, 'Recovering'],
        ['Z-011', 'Plains Zebra', 'Floodplain', 355, 'Healthy'],
        ['Z-012', 'Grevy\'s Zebra', 'Semi-arid', 400, 'Healthy'],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 18

    # --- Sheet 2: Apple ---
    ws2 = wb.create_sheet('Apple')
    headers2 = ['Variety', 'Origin', 'Season', 'Price per kg ($)', 'Stock (tons)']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FF70AD47", end_color="FF70AD47", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    data2 = [
        ['Fuji', 'Japan', 'Autumn', 3.50, 120],
        ['Gala', 'New Zealand', 'Late Summer', 2.80, 95],
        ['Granny Smith', 'Australia', 'Autumn', 2.40, 150],
        ['Honeycrisp', 'USA', 'Autumn', 4.20, 65],
        ['Pink Lady', 'Australia', 'Winter', 3.90, 80],
        ['Red Delicious', 'USA', 'Autumn', 2.10, 200],
        ['Braeburn', 'New Zealand', 'Autumn', 3.10, 110],
        ['Golden Delicious', 'France', 'Autumn', 2.60, 175],
        ['McIntosh', 'Canada', 'Early Autumn', 2.90, 88],
        ['Ambrosia', 'Canada', 'Autumn', 3.70, 55],
        ['Envy', 'New Zealand', 'Autumn', 4.50, 42],
    ]
    for r, row_data in enumerate(data2, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 16
    ws2.column_dimensions['E'].width = 14

    # --- Sheet 3: Mango ---
    ws3 = wb.create_sheet('Mango')
    headers3 = ['Cultivar', 'Region', 'Harvest Month', 'Sweetness (Brix)', 'Export Volume']
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")
        cell.font = Font(bold=True)
    data3 = [
        ['Alphonso', 'Maharashtra, India', 'April', 21.5, 45000],
        ['Kent', 'Florida, USA', 'June', 18.2, 32000],
        ['Tommy Atkins', 'Brazil', 'January', 15.8, 78000],
        ['Ataulfo', 'Mexico', 'March', 22.0, 25000],
        ['Haden', 'Florida, USA', 'May', 17.5, 18000],
        ['Keitt', 'California, USA', 'August', 16.9, 12000],
        ['Palmer', 'Brazil', 'February', 17.0, 55000],
        ['Nam Doc Mai', 'Thailand', 'April', 23.1, 30000],
        ['Carabao', 'Philippines', 'March', 20.5, 40000],
        ['Langra', 'Uttar Pradesh, India', 'June', 19.8, 22000],
        ['Chaunsa', 'Punjab, Pakistan', 'July', 21.0, 28000],
        ['Irwin', 'Japan', 'June', 18.5, 5000],
    ]
    for r, row_data in enumerate(data3, 2):
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r, column=c, value=val)
    ws3.column_dimensions['A'].width = 18
    ws3.column_dimensions['B'].width = 24
    ws3.column_dimensions['C'].width = 16
    ws3.column_dimensions['D'].width = 18
    ws3.column_dimensions['E'].width = 16

    # --- Sheet 4: Banana ---
    ws4 = wb.create_sheet('Banana')
    headers4 = ['Type', 'Country', 'Annual Production (MT)', 'Avg Price ($)', 'Organic']
    for col, h in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFED7D31", end_color="FFED7D31", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    data4 = [
        ['Cavendish', 'Ecuador', 6800000, 0.85, 'No'],
        ['Cavendish', 'Philippines', 5900000, 0.78, 'No'],
        ['Cavendish', 'Costa Rica', 2400000, 0.92, 'Yes'],
        ['Plantain', 'Colombia', 3800000, 0.65, 'No'],
        ['Red Banana', 'India', 1200000, 1.40, 'No'],
        ['Lady Finger', 'Australia', 350000, 2.10, 'Yes'],
        ['Blue Java', 'Hawaii, USA', 95000, 3.50, 'Yes'],
        ['Burro', 'Mexico', 520000, 1.15, 'No'],
        ['Manzano', 'Central America', 280000, 1.85, 'No'],
        ['Plantain', 'Ghana', 4200000, 0.55, 'No'],
        ['Cavendish', 'India', 31000000, 0.42, 'No'],
    ]
    for r, row_data in enumerate(data4, 2):
        for c, val in enumerate(row_data, 1):
            ws4.cell(row=r, column=c, value=val)
    ws4.column_dimensions['A'].width = 16
    ws4.column_dimensions['B'].width = 20
    ws4.column_dimensions['C'].width = 22
    ws4.column_dimensions['D'].width = 14
    ws4.column_dimensions['E'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet order: {wb.sheetnames}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
