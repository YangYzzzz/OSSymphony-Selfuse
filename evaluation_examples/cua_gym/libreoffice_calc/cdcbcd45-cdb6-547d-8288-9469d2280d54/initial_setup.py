"""
Initial Setup: Grade curve analysis spreadsheet with raw student scores
Task ID: calc_wf_017
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_017'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grades"

    # --- Styling ---
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # --- Headers ---
    headers = ['Student', 'Raw Score', 'Z-Score', 'Curved Score', 'Grade']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- 30 Students with realistic names and raw scores 40-95 ---
    students = [
        ("Sarah Chen", 87), ("Marcus Johnson", 72), ("Priya Patel", 91),
        ("James O'Brien", 65), ("Yuki Tanaka", 78), ("Elena Rodriguez", 83),
        ("David Kim", 55), ("Aisha Mohammed", 94), ("Tyler Brooks", 68),
        ("Mei-Lin Wu", 76), ("Carlos Gutierrez", 82), ("Hannah Fischer", 61),
        ("Raj Kapoor", 88), ("Olivia Thompson", 74), ("Andre Williams", 47),
        ("Sofia Ivanova", 90), ("Nathan Park", 71), ("Grace Okonkwo", 85),
        ("Lucas Moreau", 59), ("Fatima Al-Rashid", 79), ("Benjamin Hart", 93),
        ("Chloe Nguyen", 67), ("Omar Hassan", 81), ("Isabella Costa", 73),
        ("Ethan Clarke", 40), ("Zara Singh", 86), ("Ryan Mitchell", 70),
        ("Amara Diallo", 77), ("Daniel Kowalski", 63), ("Lily Chang", 95),
    ]

    for r, (name, score) in enumerate(students, 2):
        ws.cell(row=r, column=1, value=name).border = thin_border
        ws.cell(row=r, column=2, value=score).border = thin_border
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
        # C, D, E columns left empty (task requires agent to fill these)
        for c in range(3, 6):
            ws.cell(row=r, column=c).border = thin_border

    # --- Statistics Section (row 33+) ---
    ws.cell(row=33, column=1, value="Mean").font = Font(bold=True)
    ws.cell(row=34, column=1, value="Std Dev").font = Font(bold=True)
    ws.cell(row=35, column=1, value="Median").font = Font(bold=True)
    # B33, B34, B35 left empty - agent must fill with formulas

    # --- Frequency Distribution Table (row 37+) ---
    freq_header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    freq_header_fill = PatternFill(start_color="FF548235", end_color="FF548235", fill_type="solid")

    ws.cell(row=37, column=1, value="Grade Range").font = freq_header_font
    ws.cell(row=37, column=1).fill = freq_header_fill
    ws.cell(row=37, column=1).alignment = header_align
    ws.cell(row=37, column=1).border = thin_border

    ws.cell(row=37, column=2, value="Frequency").font = freq_header_font
    ws.cell(row=37, column=2).fill = freq_header_fill
    ws.cell(row=37, column=2).alignment = header_align
    ws.cell(row=37, column=2).border = thin_border

    bins = ["0-59", "60-69", "70-79", "80-89", "90-100"]
    for i, b in enumerate(bins):
        ws.cell(row=38 + i, column=1, value=b).border = thin_border
        ws.cell(row=38 + i, column=1).alignment = Alignment(horizontal="center")
        # Frequency column left empty - agent fills with FREQUENCY or COUNTIFS
        ws.cell(row=38 + i, column=2).border = thin_border

    # --- Column widths ---
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10

    # --- Freeze header row ---
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
