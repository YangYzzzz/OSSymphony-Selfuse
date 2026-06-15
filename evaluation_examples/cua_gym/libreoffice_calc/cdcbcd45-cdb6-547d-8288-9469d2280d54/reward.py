"""
Reward Script: Grade Curve Analysis
Task ID: calc_wf_017
Domain: libreoffice_calc
Scoring:
  Component 1: Statistics formulas (Mean/StdDev/Median) in B33:B35 — 0.20 points
  Component 2: Z-Score formulas in C2:C31 — 0.20 points
  Component 3: Curved Score formulas in D2:D31 — 0.15 points
  Component 4: Grade assignment formulas in E2:E31 — 0.15 points
  Component 5: Frequency distribution formulas in B38:B42 — 0.15 points
  Component 6: Histogram/bar chart present — 0.15 points
"""

import os
import openpyxl
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_017'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Grades' sheet must exist
    if 'Grades' not in wb.sheetnames:
        print("FAIL: 'Grades' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Grades']

    # ---------------------------------------------------------------
    # Component 1: Statistics formulas in B33, B34, B35 (0.20 points)
    # Initial state: these cells are empty. Golden state: AVERAGE, STDEV, MEDIAN.
    # ---------------------------------------------------------------
    try:
        b33 = ws['B33'].value
        b34 = ws['B34'].value
        b35 = ws['B35'].value

        # Check Mean formula (0.0667 pts)
        if b33 and isinstance(b33, str) and 'AVERAGE' in b33.upper():
            print(f"PASS: B33 contains AVERAGE formula: {b33}")
            total_score += 0.0667
        else:
            print(f"FAIL: B33 expected AVERAGE formula, found: {b33}")

        # Check StdDev formula (0.0667 pts)
        if b34 and isinstance(b34, str) and 'STDEV' in b34.upper():
            print(f"PASS: B34 contains STDEV formula: {b34}")
            total_score += 0.0667
        else:
            print(f"FAIL: B34 expected STDEV formula, found: {b34}")

        # Check Median formula (0.0666 pts)
        if b35 and isinstance(b35, str) and 'MEDIAN' in b35.upper():
            print(f"PASS: B35 contains MEDIAN formula: {b35}")
            total_score += 0.0666
        else:
            print(f"FAIL: B35 expected MEDIAN formula, found: {b35}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ---------------------------------------------------------------
    # Component 2: Z-Score formulas in C2:C31 (0.20 points)
    # Initial state: C2:C31 are empty. Golden: formula referencing B and division.
    # Z-Score = (raw - mean) / stdev
    # ---------------------------------------------------------------
    try:
        z_count = 0
        for r in range(2, 32):
            val = ws.cell(row=r, column=3).value  # column C
            if val and isinstance(val, str) and val.startswith('='):
                v_upper = val.upper().replace(' ', '')
                if 'B' in v_upper and '/' in v_upper:
                    z_count += 1
        if z_count >= 25:
            print(f"PASS: Component 2 — {z_count}/30 Z-Score formulas found (0.20 pts)")
            total_score += 0.20
        elif z_count >= 15:
            print(f"PARTIAL: Component 2 — {z_count}/30 Z-Score formulas (0.10 pts)")
            total_score += 0.10
        elif z_count > 0:
            print(f"PARTIAL: Component 2 — {z_count}/30 Z-Score formulas (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 2 — No Z-Score formulas found in C2:C31")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---------------------------------------------------------------
    # Component 3: Curved Score formulas in D2:D31 (0.15 points)
    # Initial state: D2:D31 are empty. Golden: scaling formula referencing C column.
    # ---------------------------------------------------------------
    try:
        curved_count = 0
        for r in range(2, 32):
            val = ws.cell(row=r, column=4).value  # column D
            if val and isinstance(val, str) and val.startswith('='):
                v_upper = val.upper().replace(' ', '')
                if 'C' in v_upper:
                    curved_count += 1
        if curved_count >= 25:
            print(f"PASS: Component 3 — {curved_count}/30 Curved Score formulas (0.15 pts)")
            total_score += 0.15
        elif curved_count >= 15:
            print(f"PARTIAL: Component 3 — {curved_count}/30 Curved Score formulas (0.08 pts)")
            total_score += 0.08
        elif curved_count > 0:
            print(f"PARTIAL: Component 3 — {curved_count}/30 Curved Score formulas (0.04 pts)")
            total_score += 0.04
        else:
            print(f"FAIL: Component 3 — No Curved Score formulas found in D2:D31")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ---------------------------------------------------------------
    # Component 4: Grade assignment formulas in E2:E31 (0.15 points)
    # Initial state: E2:E31 are empty. Golden: IF/VLOOKUP grade assignment.
    # ---------------------------------------------------------------
    try:
        grade_count = 0
        for r in range(2, 32):
            val = ws.cell(row=r, column=5).value  # column E
            if val and isinstance(val, str) and val.startswith('='):
                v_upper = val.upper().replace(' ', '')
                if 'IF' in v_upper or 'VLOOKUP' in v_upper or 'LOOKUP' in v_upper:
                    grade_count += 1
        if grade_count >= 25:
            print(f"PASS: Component 4 — {grade_count}/30 Grade formulas (0.15 pts)")
            total_score += 0.15
        elif grade_count >= 15:
            print(f"PARTIAL: Component 4 — {grade_count}/30 Grade formulas (0.08 pts)")
            total_score += 0.08
        elif grade_count > 0:
            print(f"PARTIAL: Component 4 — {grade_count}/30 Grade formulas (0.04 pts)")
            total_score += 0.04
        else:
            print(f"FAIL: Component 4 — No Grade formulas found in E2:E31")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ---------------------------------------------------------------
    # Component 5: Frequency distribution formulas in B38:B42 (0.15 points)
    # Initial state: B38:B42 are empty. Golden: COUNTIFS/FREQUENCY formulas.
    # ---------------------------------------------------------------
    try:
        freq_count = 0
        for r in range(38, 43):
            val = ws.cell(row=r, column=2).value  # column B
            if val and isinstance(val, str) and val.startswith('='):
                v_upper = val.upper().replace(' ', '')
                if 'COUNTIF' in v_upper or 'FREQUENCY' in v_upper:
                    freq_count += 1
        if freq_count >= 4:
            print(f"PASS: Component 5 — {freq_count}/5 Frequency formulas (0.15 pts)")
            total_score += 0.15
        elif freq_count >= 2:
            print(f"PARTIAL: Component 5 — {freq_count}/5 Frequency formulas (0.08 pts)")
            total_score += 0.08
        elif freq_count > 0:
            print(f"PARTIAL: Component 5 — {freq_count}/5 Frequency formulas (0.04 pts)")
            total_score += 0.04
        else:
            print(f"FAIL: Component 5 — No Frequency formulas found in B38:B42")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # ---------------------------------------------------------------
    # Component 6: Histogram/bar chart present (0.15 points)
    # Initial state: no charts. Golden: 1 bar/column chart.
    # ---------------------------------------------------------------
    try:
        charts = ws._charts
        if len(charts) >= 1:
            chart = charts[0]
            chart_type = getattr(chart, 'type', None)
            if chart_type in ('col', 'bar'):
                print(f"PASS: Component 6 — Bar/column chart found (type={chart_type}) (0.15 pts)")
                total_score += 0.15
            else:
                print(f"PARTIAL: Component 6 — Chart found but type={chart_type}, expected bar/col (0.08 pts)")
                total_score += 0.08
        else:
            print(f"FAIL: Component 6 — No charts found in Grades sheet")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
