"""
Initial Setup: Rename sheet task - create workbook with Sheet1, Sheet2, Sheet3
Task ID: calc_ps_048
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_048'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: Employee data ---
    ws1 = wb.active
    ws1.title = 'Sheet1'

    headers = ['Name', 'Department', 'Salary', 'Start Date', 'Email']
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center")

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_alignment

    data = [
        ['Sarah Chen', 'Engineering', 95000, '2022-03-15', 'sarah.chen@acmecorp.com'],
        ['Marcus Johnson', 'Marketing', 72000, '2021-06-01', 'marcus.j@acmecorp.com'],
        ['Priya Patel', 'Engineering', 88500, '2023-01-10', 'priya.patel@acmecorp.com'],
        ['David Kim', 'Finance', 81000, '2020-09-22', 'david.kim@acmecorp.com'],
        ['Elena Rodriguez', 'Human Resources', 68000, '2022-11-05', 'elena.r@acmecorp.com'],
        ['James Wright', 'Engineering', 102000, '2019-04-18', 'james.wright@acmecorp.com'],
        ['Aisha Mohammed', 'Marketing', 75500, '2023-07-20', 'aisha.m@acmecorp.com'],
        ['Thomas Berg', 'Finance', 79000, '2021-02-14', 'thomas.berg@acmecorp.com'],
        ['Lisa Nakamura', 'Engineering', 91000, '2022-08-30', 'lisa.n@acmecorp.com'],
        ['Carlos Mendez', 'Operations', 66500, '2023-05-12', 'carlos.mendez@acmecorp.com'],
        ['Rachel Foster', 'Human Resources', 71000, '2020-12-01', 'rachel.f@acmecorp.com'],
        ['Kevin O\'Brien', 'Operations', 69500, '2021-10-08', 'kevin.obrien@acmecorp.com'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Set salary column as currency format
    for r in range(2, len(data) + 2):
        ws1.cell(row=r, column=3).number_format = '$#,##0.00'

    # Set column widths for readability
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 14
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 28

    # --- Sheet2: Department Summary (some reference data) ---
    ws2 = wb.create_sheet('Sheet2')
    ws2.cell(row=1, column=1, value='Department')
    ws2.cell(row=1, column=2, value='Budget')
    ws2.cell(row=1, column=3, value='Headcount Target')

    dept_data = [
        ['Engineering', 500000, 15],
        ['Marketing', 250000, 8],
        ['Finance', 300000, 10],
        ['Human Resources', 200000, 6],
        ['Operations', 180000, 7],
    ]
    for r, row_data in enumerate(dept_data, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    # --- Sheet3: Notes ---
    ws3 = wb.create_sheet('Sheet3')
    ws3.cell(row=1, column=1, value='Notes')
    ws3.cell(row=2, column=1, value='Q1 2025 hiring plan under review.')
    ws3.cell(row=3, column=1, value='Budget approval pending for Engineering expansion.')
    ws3.column_dimensions['A'].width = 50

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
