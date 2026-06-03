"""
Reward Script: Rename sheet 'Sheet1' to 'Employee Records'
Task ID: calc_ps_048
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Sheet 'Employee Records' exists
  Component 2 (0.3): Sheet 'Sheet1' does NOT exist (rename, not copy)
  Component 3 (0.3): Data integrity - 'Employee Records' has original employee data
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_048'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    sheet_names = wb.sheetnames
    print(f"INFO: Found sheets: {sheet_names}")

    # Component 1: Sheet 'Employee Records' exists (0.4 points)
    # This checks the core rename result - the new name must be present
    try:
        if 'Employee Records' in sheet_names:
            print(f"PASS: Component 1 - 'Employee Records' sheet exists (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 - 'Employee Records' sheet not found in {sheet_names}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Sheet 'Sheet1' does NOT exist (0.3 points)
    # This ensures it was a rename (not a copy/add). The old name must be gone.
    try:
        if 'Sheet1' not in sheet_names:
            print(f"PASS: Component 2 - 'Sheet1' no longer exists (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 - 'Sheet1' still exists (should have been renamed, not copied)")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Data integrity - 'Employee Records' has the original employee data (0.3 points)
    # Verify key data cells to ensure data was preserved during rename
    try:
        if 'Employee Records' in sheet_names:
            ws = wb['Employee Records']
            # Check header row
            headers_ok = (
                ws['A1'].value == 'Name' and
                ws['B1'].value == 'Department' and
                ws['C1'].value == 'Salary'
            )
            # Check a few data values
            data_ok = (
                ws['A2'].value == 'Sarah Chen' and
                ws['B2'].value == 'Engineering' and
                ws['C2'].value == 95000
            )
            # Check row count is reasonable (13 rows in original)
            row_count = ws.max_row
            size_ok = row_count >= 10

            if headers_ok and data_ok and size_ok:
                print(f"PASS: Component 3 - Data integrity verified (headers, data, {row_count} rows) (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 3 - Data integrity issues: headers_ok={headers_ok}, data_ok={data_ok}, size_ok={size_ok} (rows={row_count})")
        else:
            print(f"FAIL: Component 3 - Cannot check data, 'Employee Records' sheet missing")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
