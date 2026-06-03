"""
Initial Setup: Daily Operations Standup Report
Task ID: calc_ops_operations_daily_standup_075
Domain: libreoffice_calc

Creates:
- Sheet 'OrderData': Date, Orders Received, Orders Dispatched, Orders On Time, Picking Errors (30 days)
- Sheet 'BacklogData': Date, Outstanding Orders (30 days)
- Sheet 'DailyReport': BLANK (to be populated by agent)
"""

import os
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_operations_daily_standup_075'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ #
    # Sheet 1: OrderData                                                   #
    # ------------------------------------------------------------------ #
    ws_order = wb.active
    ws_order.title = 'OrderData'

    # Headers
    order_headers = ['Date', 'Orders Received', 'Orders Dispatched', 'Orders On Time', 'Picking Errors']
    for col, h in enumerate(order_headers, 1):
        cell = ws_order.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # 30 days of data ending with yesterday
    today = date.today()
    yesterday = today - timedelta(days=1)

    # Realistic daily operations data
    import random
    random.seed(42)  # reproducible data

    order_rows = []
    for i in range(29, -1, -1):
        day = today - timedelta(days=i)
        received = random.randint(185, 235)
        dispatched = random.randint(int(received * 0.88), int(received * 0.98))
        on_time_pct = random.uniform(0.90, 0.99)
        on_time = int(dispatched * on_time_pct)
        errors = random.randint(0, int(dispatched * 0.012))
        order_rows.append([day, received, dispatched, on_time, errors])

    for r, row_data in enumerate(order_rows, 2):
        ws_order.cell(row=r, column=1, value=row_data[0]).number_format = 'yyyy-mm-dd'
        ws_order.cell(row=r, column=2, value=row_data[1])
        ws_order.cell(row=r, column=3, value=row_data[2])
        ws_order.cell(row=r, column=4, value=row_data[3])
        ws_order.cell(row=r, column=5, value=row_data[4])

    # Column widths
    ws_order.column_dimensions['A'].width = 14
    ws_order.column_dimensions['B'].width = 18
    ws_order.column_dimensions['C'].width = 20
    ws_order.column_dimensions['D'].width = 18
    ws_order.column_dimensions['E'].width = 16

    # ------------------------------------------------------------------ #
    # Sheet 2: BacklogData                                                 #
    # ------------------------------------------------------------------ #
    ws_backlog = wb.create_sheet('BacklogData')

    backlog_headers = ['Date', 'Outstanding Orders']
    for col, h in enumerate(backlog_headers, 1):
        cell = ws_backlog.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for i in range(29, -1, -1):
        day = today - timedelta(days=i)
        outstanding = random.randint(12, 68)
        row_idx = 30 - i + 1
        ws_backlog.cell(row=row_idx, column=1, value=day).number_format = 'yyyy-mm-dd'
        ws_backlog.cell(row=row_idx, column=2, value=outstanding)

    ws_backlog.column_dimensions['A'].width = 14
    ws_backlog.column_dimensions['B'].width = 20

    # ------------------------------------------------------------------ #
    # Sheet 3: DailyReport (BLANK — agent must populate)                  #
    # ------------------------------------------------------------------ #
    ws_report = wb.create_sheet('DailyReport')
    # Intentionally left blank

    # Column widths pre-set so it looks ready for content
    ws_report.column_dimensions['A'].width = 28
    ws_report.column_dimensions['B'].width = 18
    ws_report.column_dimensions['C'].width = 20
    ws_report.column_dimensions['D'].width = 16
    ws_report.column_dimensions['E'].width = 12
    ws_report.column_dimensions['F'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheets: {wb.sheetnames}')
    print(f'OrderData rows: {ws_order.max_row - 1} data rows')
    print(f'BacklogData rows: {ws_backlog.max_row - 1} data rows')
    print(f'DailyReport: BLANK (ready for agent)')


create_initial()
