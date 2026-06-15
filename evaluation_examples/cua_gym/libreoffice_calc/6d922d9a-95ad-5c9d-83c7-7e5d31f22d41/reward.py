"""
Reward Script: Daily Operations Standup Report Template
Task ID: calc_ops_operations_daily_standup_075
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: DailyReport title row — A1 has title text, merged A1:D1, date formulas in B2/B3 (0.15)
  Component 2: KPI header row — A5:E5 populated with correct header labels, bold (0.15)
  Component 3: KPI rows structure — Rows 6-9 have correct KPI names and target values (0.20)
  Component 4: KPI actuals — C6:C9 have INDEX/MATCH formulas referencing OrderData/BacklogData (0.20)
  Component 5: Freeze panes at A5 (row 4 freeze) (0.10)
  Component 6: Print settings — landscape orientation, fit to 1 page, print area A1:F12 (0.10)
  Component 7: Conditional formatting on E6:E9 Status column (0.10)
  Total: 1.0
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_ops_operations_daily_standup_075'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: DailyReport sheet must exist
    if 'DailyReport' not in wb.sheetnames:
        print("CRITICAL: 'DailyReport' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['DailyReport']

    # Component 1: Title row structure (0.15 points)
    # A1 must have the title text, merged to at least D1 or F1,
    # B2 must have =TODAY() formula, B3 must have =TODAY()-1 formula
    try:
        a1_val = ws['A1'].value
        b2_val = ws['B2'].value
        b3_val = ws['B3'].value
        a2_val = ws['A2'].value
        a3_val = ws['A3'].value

        title_ok = (
            a1_val is not None and
            'DAILY' in str(a1_val).upper() and
            'STANDUP' in str(a1_val).upper()
        )

        # Check A1 is part of a merge (the merge range spans at least to D1)
        merge_ok = False
        for merge_range in ws.merged_cells.ranges:
            if (merge_range.min_row == 1 and merge_range.max_row >= 1 and
                    merge_range.min_col == 1 and merge_range.max_col >= 4):
                merge_ok = True
                break

        b2_formula = isinstance(b2_val, str) and 'TODAY()' in b2_val.upper()
        b3_formula = isinstance(b3_val, str) and 'TODAY()' in b3_val.upper() and '-1' in b3_val

        a2_label = a2_val is not None and 'DATE' in str(a2_val).upper()
        a3_label = a3_val is not None and ('PERIOD' in str(a3_val).upper() or 'REPORT' in str(a3_val).upper())

        if title_ok and merge_ok and b2_formula and b3_formula:
            print(f"PASS: Component 1 — Title '{a1_val}' present, merged, date formulas in B2/B3 (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — title_ok={title_ok}, merge_ok={merge_ok}, b2_formula={b2_formula}, b3_formula={b3_formula}")
            print(f"      A1={repr(a1_val)}, B2={repr(b2_val)}, B3={repr(b3_val)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: KPI header row (0.15 points)
    # Row 5 must have correct headers in A5:E5, with bold formatting
    try:
        expected_headers = {
            1: 'KPI',
            2: 'Target',
            3: 'Yesterday Actual',
            4: 'Achievement',  # partial match for "Achievement %"
            5: 'Status'
        }
        headers_ok = True
        headers_bold = True
        header_details = []

        for col, expected in expected_headers.items():
            cell = ws.cell(row=5, column=col)
            val = cell.value
            if val is None or expected.upper() not in str(val).upper():
                headers_ok = False
                header_details.append(f"col{col} expected '{expected}', got {repr(val)}")
            if not cell.font.bold:
                headers_bold = False
                header_details.append(f"col{col} not bold")

        if headers_ok and headers_bold:
            print(f"PASS: Component 2 — KPI headers in row 5, all bold (0.15 pts)")
            total_score += 0.15
        elif headers_ok:
            # Partial credit: headers present but not bold
            print(f"PARTIAL: Component 2 — KPI headers correct but not bold (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 2 — headers: {'; '.join(header_details)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: KPI row names and target values (0.20 points)
    # Rows 6-9 must have correct KPI names (col A) and target values (col B)
    try:
        expected_kpi_names = {
            6: 'Orders Dispatched',
            7: 'On-Time',
            8: 'Picking Accuracy',
            9: 'Backlog'
        }
        expected_targets = {
            6: 200,
            7: 0.95,
            8: 0.995,
            9: 0  # target 0 for backlog
        }

        names_ok = 0
        targets_ok = 0

        for row, expected_name in expected_kpi_names.items():
            a_val = ws.cell(row=row, column=1).value
            b_val = ws.cell(row=row, column=2).value
            expected_target = expected_targets[row]

            if a_val is not None and expected_name.upper() in str(a_val).upper():
                names_ok += 1
            else:
                print(f"  FAIL: Row {row} KPI name: expected containing '{expected_name}', got {repr(a_val)}")

            # Compare target values
            if b_val is not None:
                try:
                    bv = float(b_val)
                    if abs(bv - expected_target) < 0.001:
                        targets_ok += 1
                    else:
                        print(f"  FAIL: Row {row} target: expected {expected_target}, got {bv}")
                except (ValueError, TypeError):
                    print(f"  FAIL: Row {row} target not numeric: {repr(b_val)}")
            else:
                print(f"  FAIL: Row {row} target is None")

        total_kpi_checks = 4
        if names_ok == total_kpi_checks and targets_ok == total_kpi_checks:
            print(f"PASS: Component 3 — All 4 KPI rows have correct names and targets (0.20 pts)")
            total_score += 0.20
        elif names_ok >= 3 and targets_ok >= 3:
            print(f"PARTIAL: Component 3 — {names_ok}/4 names, {targets_ok}/4 targets correct (0.12 pts)")
            total_score += 0.12
        elif names_ok >= 2 or targets_ok >= 2:
            print(f"PARTIAL: Component 3 — {names_ok}/4 names, {targets_ok}/4 targets correct (0.06 pts)")
            total_score += 0.06
        else:
            print(f"FAIL: Component 3 — {names_ok}/4 names, {targets_ok}/4 targets correct")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: KPI actuals use INDEX/MATCH formulas referencing data sheets (0.20 points)
    # C6:C9 must contain formulas using INDEX/MATCH that pull from OrderData or BacklogData
    try:
        formulas_ok = 0
        formula_details = []

        for row in range(6, 10):
            c_val = ws.cell(row=row, column=3).value
            if c_val is None:
                formula_details.append(f"Row {row}: C is None")
                continue
            val_str = str(c_val).upper()

            # Must be a formula
            if not val_str.startswith('='):
                formula_details.append(f"Row {row}: not a formula: {repr(c_val)}")
                continue

            # Check it references the right data source
            if row in (6, 7, 8):
                # Should reference OrderData
                if 'ORDERDATA' in val_str and ('INDEX' in val_str or 'VLOOKUP' in val_str or 'MATCH' in val_str):
                    formulas_ok += 1
                else:
                    formula_details.append(f"Row {row}: missing INDEX/MATCH on OrderData: {repr(c_val)}")
            elif row == 9:
                # Should reference BacklogData
                if 'BACKLOGDATA' in val_str and ('INDEX' in val_str or 'VLOOKUP' in val_str or 'MATCH' in val_str):
                    formulas_ok += 1
                else:
                    formula_details.append(f"Row {row}: missing INDEX/MATCH on BacklogData: {repr(c_val)}")

        if formulas_ok == 4:
            print(f"PASS: Component 4 — All 4 KPI actual formulas reference correct data sheets (0.20 pts)")
            total_score += 0.20
        elif formulas_ok >= 3:
            print(f"PARTIAL: Component 4 — {formulas_ok}/4 formulas reference correct sheets (0.13 pts)")
            total_score += 0.13
        elif formulas_ok >= 2:
            print(f"PARTIAL: Component 4 — {formulas_ok}/4 formulas reference correct sheets (0.08 pts)")
            total_score += 0.08
        elif formulas_ok >= 1:
            print(f"PARTIAL: Component 4 — {formulas_ok}/4 formulas reference correct sheets (0.04 pts)")
            total_score += 0.04
        else:
            print(f"FAIL: Component 4 — no valid INDEX/MATCH formulas found; {'; '.join(formula_details)}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Freeze panes at A5 (0.10 points)
    # Freeze panes must be set to A5 (freezes rows 1-4 so KPI header row 5 stays visible)
    try:
        freeze = ws.freeze_panes
        if freeze == 'A5':
            print(f"PASS: Component 5 — Freeze panes set to A5 (0.10 pts)")
            total_score += 0.10
        elif freeze is not None and str(freeze).startswith('A') and int(str(freeze)[1:]) >= 2:
            # Some freeze was applied, partial credit
            print(f"PARTIAL: Component 5 — Freeze panes = {freeze} (expected A5) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — Freeze panes = {repr(freeze)}, expected 'A5'")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Print settings — landscape, fit to 1 page, print area (0.10 points)
    try:
        ps = ws.page_setup
        orientation_ok = ps.orientation == 'landscape'
        fit_ok = (ps.fitToWidth == 1 and ps.fitToHeight == 1)
        print_area = ws.print_area
        # print_area may be like "'DailyReport'!$A$1:$F$12" or "$A$1:$F$12"
        area_ok = print_area is not None and 'A1' in str(print_area).upper().replace('$', '') and \
                  'F12' in str(print_area).upper().replace('$', '')

        if orientation_ok and fit_ok and area_ok:
            print(f"PASS: Component 6 — Landscape, fit to 1 page, print area covers A1:F12 (0.10 pts)")
            total_score += 0.10
        elif orientation_ok and fit_ok:
            print(f"PARTIAL: Component 6 — Landscape + fit-to-page set, but print area missing/wrong: {print_area} (0.07 pts)")
            total_score += 0.07
        elif orientation_ok or fit_ok:
            print(f"PARTIAL: Component 6 — orientation_ok={orientation_ok}, fit_ok={fit_ok}, area_ok={area_ok} (0.04 pts)")
            total_score += 0.04
        else:
            print(f"FAIL: Component 6 — orientation={ps.orientation}, fitToWidth={ps.fitToWidth}, fitToHeight={ps.fitToHeight}, print_area={print_area}")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: Conditional formatting on Status column E6:E9 (0.10 points)
    # Must have at least 2 conditional formatting rules on the E6:E9 range
    try:
        cf_rules = ws.conditional_formatting
        cf_on_status = False
        rule_count = 0

        for cf_range in cf_rules:
            range_str = str(cf_range).upper()
            # Check if the conditional formatting applies to column E rows 6-9
            if 'E6' in range_str or 'E:E' in range_str:
                rules = cf_rules[cf_range]
                rule_count += len(rules)
                if len(rules) >= 1:
                    cf_on_status = True

        if cf_on_status and rule_count >= 2:
            print(f"PASS: Component 7 — Conditional formatting on Status column with {rule_count} rules (0.10 pts)")
            total_score += 0.10
        elif cf_on_status:
            print(f"PARTIAL: Component 7 — Conditional formatting found but only {rule_count} rule(s) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 7 — No conditional formatting found on Status column E6:E9")
            # Check if there's any CF at all
            all_cf = list(cf_rules)
            if all_cf:
                print(f"  (Found CF on: {[str(x) for x in all_cf]})")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
