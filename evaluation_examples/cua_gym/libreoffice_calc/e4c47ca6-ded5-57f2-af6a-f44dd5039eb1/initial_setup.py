"""
Initial Setup: Create Project_Assignments spreadsheet with tasks and status values
Task ID: calc_gcv_058
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_058'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: Project Assignments ---
    ws1 = wb.active
    ws1.title = 'Sheet1'

    headers = ['Task', 'Priority', 'Assignee', 'Status']
    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font

    # Column widths for readability
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 15

    # 39 rows of realistic project task data
    tasks = [
        ['Set up CI/CD pipeline for backend services', 'High', 'Sarah Chen'],
        ['Design landing page mockups', 'Medium', 'Marcus Johnson'],
        ['Implement user authentication module', 'High', 'Priya Sharma'],
        ['Write unit tests for payment gateway', 'High', 'David Kim'],
        ['Configure load balancer settings', 'Medium', 'Elena Rodriguez'],
        ['Create API documentation for v2 endpoints', 'Low', 'James Wilson'],
        ['Optimize database query performance', 'High', 'Mei Lin'],
        ['Set up monitoring dashboards in Grafana', 'Medium', 'Carlos Rivera'],
        ['Migrate legacy user data to new schema', 'High', 'Aisha Patel'],
        ['Implement email notification service', 'Medium', 'Tom Anderson'],
        ['Review security audit findings', 'High', 'Rachel Green'],
        ['Build admin panel for content management', 'Medium', 'Kevin Park'],
        ['Deploy staging environment on AWS', 'High', 'Fatima Al-Hassan'],
        ['Create onboarding tutorial for new users', 'Low', 'Brian O\'Neill'],
        ['Fix memory leak in background workers', 'High', 'Yuki Tanaka'],
        ['Design mobile-responsive navigation', 'Medium', 'Sofia Martinez'],
        ['Implement rate limiting for public API', 'High', 'Alex Thompson'],
        ['Set up automated backup procedures', 'Medium', 'Nina Petrov'],
        ['Build real-time chat feature', 'High', 'Lucas Ferreira'],
        ['Conduct performance benchmarks', 'Medium', 'Hannah Lee'],
        ['Update SSL certificates for production', 'High', 'Omar Hussain'],
        ['Create data export functionality', 'Low', 'Emily Watson'],
        ['Implement two-factor authentication', 'High', 'Raj Gupta'],
        ['Design dashboard analytics widgets', 'Medium', 'Clara Hoffmann'],
        ['Set up error tracking with Sentry', 'Medium', 'Michael Brown'],
        ['Build file upload and processing pipeline', 'High', 'Wei Zhang'],
        ['Write integration tests for checkout flow', 'High', 'Jessica Taylor'],
        ['Configure CDN for static assets', 'Medium', 'Daniel Park'],
        ['Implement webhook delivery system', 'Medium', 'Lena Johansson'],
        ['Create customer feedback survey form', 'Low', 'Ryan Murphy'],
        ['Optimize image compression pipeline', 'Medium', 'Ananya Singh'],
        ['Set up blue-green deployment strategy', 'High', 'Patrick O\'Connor'],
        ['Build notification preferences panel', 'Low', 'Maria Santos'],
        ['Implement search indexing with Elasticsearch', 'High', 'Chris Evans'],
        ['Design print-friendly invoice template', 'Low', 'Laura Schmidt'],
        ['Create API versioning strategy document', 'Medium', 'Ali Reza'],
        ['Build user activity audit log', 'High', 'Natalie Wright'],
        ['Implement password reset flow', 'Medium', 'Jorge Hernandez'],
        ['Set up log aggregation with ELK stack', 'High', 'Ingrid Bergstrom'],
    ]

    for r, row_data in enumerate(tasks, 2):
        ws1.cell(row=r, column=1, value=row_data[0])  # Task
        ws1.cell(row=r, column=2, value=row_data[1])  # Priority
        ws1.cell(row=r, column=3, value=row_data[2])  # Assignee
        # Column D (Status) left EMPTY - no data validation

    # --- Sheet2: Status Values ---
    ws2 = wb.create_sheet('Sheet2')
    status_values = [
        'Not Started',
        'Planning',
        'In Progress',
        'Review',
        'Testing',
        'Approved',
        'Deployed',
        'Closed',
    ]
    for r, status in enumerate(status_values, 1):
        ws2.cell(row=r, column=1, value=status)

    ws2.column_dimensions['A'].width = 15

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
