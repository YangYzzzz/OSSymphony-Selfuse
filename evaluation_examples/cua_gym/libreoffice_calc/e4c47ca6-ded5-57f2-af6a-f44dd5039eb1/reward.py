"""
Reward Script: Create dropdown list in D2:D40 referencing Sheet2.A1:A8
Task ID: calc_gcv_058
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Data validation exists on Sheet1 of type 'list'
  Component 2 (0.3): Data validation formula references Sheet2.$A$1:$A$8
  Component 3 (0.3): Data validation applies to range D2:D40
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_058'


def persist_app_state(domain):
    """Save any unsaved GUI state before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            import time
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sheet1 and Sheet2 must exist
    if 'Sheet1' not in wb.sheetnames or 'Sheet2' not in wb.sheetnames:
        print(f"FAIL: Required sheets not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Sheet1']
    validations = ws.data_validations.dataValidation

    # Component 1: A data validation of type 'list' exists on Sheet1 (0.4 points)
    try:
        list_validations = [dv for dv in validations if dv.type == 'list']
        if len(list_validations) >= 1:
            print(f"PASS: Component 1 — Found {len(list_validations)} list-type data validation(s) on Sheet1 (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — No list-type data validation found. Total validations: {len(validations)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: The data validation formula references Sheet2.$A$1:$A$8 (0.3 points)
    try:
        formula_match = False
        for dv in validations:
            if dv.type == 'list' and dv.formula1:
                formula = str(dv.formula1).strip().replace("'", "")
                # Normalize: remove quotes, spaces, and compare case-insensitively
                normalized = formula.upper().replace(" ", "")
                # Accept variants: Sheet2!$A$1:$A$8, Sheet2.$A$1:$A$8, etc.
                if "SHEET2" in normalized and "$A$1:$A$8" in normalized:
                    formula_match = True
                    print(f"PASS: Component 2 — Formula references Sheet2.$A$1:$A$8 (found: {dv.formula1}) (0.3 pts)")
                    break
        if formula_match:
            total_score += 0.3
        else:
            formulas = [f"type={dv.type}, formula1={dv.formula1}" for dv in validations]
            print(f"FAIL: Component 2 — No validation references Sheet2.$A$1:$A$8. Found: {formulas}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Data validation applies to range D2:D40 (0.3 points)
    try:
        range_match = False
        for dv in validations:
            if dv.type == 'list':
                sqref_str = str(dv.sqref).upper().replace(" ", "")
                # The sqref could be "D2:D40" or equivalent
                if "D2:D40" in sqref_str:
                    range_match = True
                    print(f"PASS: Component 3 — Validation applies to D2:D40 (found: {dv.sqref}) (0.3 pts)")
                    break
                # Also check if it covers D2:D40 via multiple ranges or a superset
                # Parse individual cell ranges from sqref
                from openpyxl.utils import range_boundaries
                try:
                    for rng in str(dv.sqref).split():
                        min_col, min_row, max_col, max_row = range_boundaries(rng)
                        # D = column 4, check if range covers D2:D40
                        if min_col == 4 and max_col == 4 and min_row <= 2 and max_row >= 40:
                            range_match = True
                            print(f"PASS: Component 3 — Validation covers D2:D40 via range {rng} (0.3 pts)")
                            break
                except Exception:
                    pass
                if range_match:
                    break
        if range_match:
            total_score += 0.3
        else:
            sqrefs = [f"sqref={dv.sqref}" for dv in validations if dv.type == 'list']
            print(f"FAIL: Component 3 — No validation covers D2:D40. Found: {sqrefs}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
