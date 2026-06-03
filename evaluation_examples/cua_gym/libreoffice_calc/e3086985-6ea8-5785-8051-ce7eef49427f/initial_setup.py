"""
Initial Setup: RACI Matrix Project Tracking
Task ID: calc_ops_project_tracking_raci_015
Domain: libreoffice_calc

Creates an initial RACI matrix spreadsheet with:
- Sheet 'RACI' with task names in column A and team member names as headers in B1:H1
- 25 task rows (rows 2-26) with task names in column A, empty RACI assignment cells (B2:H26)
- Column I is empty (no header, no formulas, no validation)
- No data validation, no conditional formatting applied yet
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_project_tracking_raci_015'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'RACI'

    # --- Row 1: Headers ---
    # Column A: 'Task', Columns B-H: 7 team member names
    team_members = [
        'Alice Chen',
        'Marcus Webb',
        'Priya Sharma',
        'Daniel Torres',
        'Yuki Nakamura',
        'Sofia Reyes',
        'Ethan Brooks'
    ]

    ws['A1'] = 'Task'
    for col_idx, name in enumerate(team_members, start=2):
        ws.cell(row=1, column=col_idx, value=name)

    # Style header row: bold, light blue background
    header_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    header_font = Font(bold=True)
    for col in range(1, 9):  # A through H
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Rows 2-26: 25 realistic project tasks ---
    tasks = [
        'Define project scope and objectives',
        'Develop project charter',
        'Stakeholder identification and analysis',
        'Requirements gathering and documentation',
        'Risk assessment and mitigation planning',
        'Resource allocation and team assignment',
        'Project timeline and milestone planning',
        'Budget estimation and approval',
        'Technology stack selection',
        'System architecture design',
        'UI/UX wireframe development',
        'Database schema design',
        'Backend API development',
        'Frontend interface implementation',
        'Integration testing and QA',
        'User acceptance testing (UAT)',
        'Performance optimization and tuning',
        'Security audit and vulnerability assessment',
        'Documentation and user manual creation',
        'Training program development',
        'Deployment pipeline setup',
        'Production environment configuration',
        'Go-live coordination and launch',
        'Post-launch monitoring and support',
        'Project retrospective and closeout',
    ]

    for row_idx, task_name in enumerate(tasks, start=2):
        ws.cell(row=row_idx, column=1, value=task_name)
        # Columns B-H (columns 2-8) are intentionally left empty

    # Set a reasonable default column width for column A (not the full 200px yet)
    ws.column_dimensions['A'].width = 20

    # Set reasonable column widths for team member columns
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col_letter].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: RACI')
    print(f'  Headers: Task + {len(team_members)} team members in B1:H1')
    print(f'  Task rows: {len(tasks)} rows (rows 2-26)')
    print(f'  RACI cells B2:H26: empty (no assignments)')
    print(f'  Column I: empty (no header, no formulas)')
    print(f'  Data validation: NONE')
    print(f'  Conditional formatting: NONE')


create_initial()
