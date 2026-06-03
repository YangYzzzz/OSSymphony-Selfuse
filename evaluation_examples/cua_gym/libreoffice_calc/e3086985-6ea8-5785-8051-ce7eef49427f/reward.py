"""
Reward Script: Create RACI matrix with data validation, validation check column, and conditional formatting
Task ID: calc_ops_project_tracking_raci_015
Domain: libreoffice_calc
Scoring:
  - Component 1: Data validation on B2:H26 allowing only R,A,C,I  (0.30 pts)
  - Component 2: 'Has R?' header (I1) and COUNTIF-based IF formulas in I2:I26  (0.35 pts)
  - Component 3: Conditional formatting on I2:I26 (red for MISSING R, green for Yes)  (0.20 pts)
  - Component 4: Freeze panes set to freeze column A (B1)  (0.10 pts)
  - Component 5: Column A width >= 25 (at least 200px equivalent)  (0.05 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_project_tracking_raci_015'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify RACI sheet exists (precondition gate)
    if 'RACI' not in wb.sheetnames:
        print("CRITICAL: Sheet 'RACI' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['RACI']

    # Component 1: Data validation on B2:H26 allowing only R, A, C, I (0.30 points)
    # Task requires: "Set up data validation on all assignment cells so users can only enter R, A, C, or I"
    # Initial file has NO data validation; golden file has type=list with "R,A,C,I" on B2:H26
    try:
        validations = ws.data_validations.dataValidation
        # Collect matching validations
        list_dvs = [dv for dv in validations if dv.type == 'list']
        raci_dvs = [
            dv for dv in list_dvs
            if dv.formula1 and all(
                letter in str(dv.formula1).upper().replace(' ', '')
                for letter in ['R', 'A', 'C', 'I']
            )
        ]
        ranged_dvs = [
            dv for dv in raci_dvs
            if 'B2' in str(dv.sqref) or 'B2:H26' in str(dv.sqref)
        ]

        if ranged_dvs:
            print(f"PASS: Component 1 — Data validation found on B2:H26 with R,A,C,I list (0.30 pts)")
            total_score += 0.30
        elif raci_dvs:
            print(f"FAIL: Component 1 — Data validation has correct values but wrong range. sqref={raci_dvs[0].sqref}")
        elif list_dvs:
            print(f"FAIL: Component 1 — Data validation found but incorrect values. formula1={list_dvs[0].formula1}")
        else:
            print(f"FAIL: Component 1 — No list-type data validation found on assignment cells")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: 'Has R?' header in I1 and COUNTIF-based IF formulas in I2:I26 (0.35 points)
    # Task requires: I1 header "Has R?", I2:I26 formula checking if any cell in row has 'R'
    # The expected formula pattern: =IF(COUNTIF(Bx:Hx,"R")>0,"Yes","MISSING R")
    try:
        i1_value = ws.cell(row=1, column=9).value
        header_ok = i1_value is not None and str(i1_value).strip().upper() == 'HAS R?'

        # Check formulas in I2:I26
        formulas_ok = 0
        total_rows = 25  # rows 2 through 26

        for row in range(2, 27):
            cell_val = ws.cell(row=row, column=9).value
            if cell_val is not None:
                val_str = str(cell_val).upper().replace(' ', '')
                # Must contain COUNTIF and reference the corresponding Bx:Hx range
                row_str = str(row)
                has_countif = 'COUNTIF' in val_str
                has_row_ref = f'B{row_str}:H{row_str}' in val_str.replace(' ', '')
                has_yes_missing = 'YES' in val_str and 'MISSINGR' in val_str
                if has_countif and has_row_ref and has_yes_missing:
                    formulas_ok += 1

        formula_ratio = formulas_ok / total_rows

        if header_ok and formula_ratio >= 0.96:  # At least 24/25 correct
            print(f"PASS: Component 2 — I1='Has R?', {formulas_ok}/{total_rows} COUNTIF formulas correct (0.35 pts)")
            total_score += 0.35
        elif header_ok and formula_ratio >= 0.5:
            # Partial credit if header is right but some formulas are missing or wrong
            partial = 0.35 * 0.5
            print(f"PARTIAL: Component 2 — I1='Has R?', but only {formulas_ok}/{total_rows} formulas correct ({partial:.2f} pts)")
            total_score += partial
        elif not header_ok:
            print(f"FAIL: Component 2 — I1 header is '{i1_value}', expected 'Has R?'")
        else:
            print(f"FAIL: Component 2 — {formulas_ok}/{total_rows} COUNTIF formulas found/correct")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Conditional formatting on I2:I26 (0.20 points)
    # Task requires: red fill for 'MISSING R', green fill for 'Yes'
    # Golden file has: expression rules on I2:I26 with FFFF0000 (red) and FF00B050 (green)
    try:
        red_rules_found = []
        green_rules_found = []

        for cf_range in ws.conditional_formatting:
            cf_range_str = str(cf_range)
            # Check if this CF applies to column I area (I2 should be in range)
            if 'I' not in cf_range_str:
                continue
            rules_list = cf_range.rules if hasattr(cf_range, 'rules') else cf_range
            if not isinstance(rules_list, list):
                continue
            for rule in rules_list:
                rule_formula = str(getattr(rule, 'formula', ''))
                has_fill = hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill and rule.dxf.fill.fgColor
                if not has_fill:
                    continue
                fill_color = rule.dxf.fill.fgColor.rgb
                if not fill_color or len(fill_color) < 8:
                    continue
                try:
                    r_val = int(fill_color[2:4], 16)
                    g_val = int(fill_color[4:6], 16)
                    b_val = int(fill_color[6:8], 16)
                except ValueError:
                    continue
                # Red rule for MISSING R
                if 'MISSING' in rule_formula.upper():
                    if r_val > 150 and g_val < 100 and b_val < 100:
                        red_rules_found.append(fill_color)
                # Green rule for Yes
                if '"YES"' in rule_formula.upper() or "'YES'" in rule_formula.upper():
                    if g_val > 100 and r_val < 100:
                        green_rules_found.append(fill_color)

        if red_rules_found and green_rules_found:
            print(f"PASS: Component 3 — CF has red (MISSING R, color={red_rules_found[0]}) and green (Yes, color={green_rules_found[0]}) rules on I column (0.20 pts)")
            total_score += 0.20
        elif red_rules_found or green_rules_found:
            partial = 0.10
            print(f"PARTIAL: Component 3 — CF found but missing one color rule. red={bool(red_rules_found)}, green={bool(green_rules_found)} ({partial} pts)")
            total_score += partial
        else:
            # Check if any CF exists at all on I column
            any_i_cf = any('I' in str(cf_range) for cf_range in ws.conditional_formatting)
            if any_i_cf:
                print(f"FAIL: Component 3 — CF found on I column but no matching color rules detected (expected red/green)")
            else:
                print(f"FAIL: Component 3 — No conditional formatting found on column I (I2:I26)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Freeze panes set to freeze column A (B1) (0.10 points)
    # Task requires: "freeze column A"
    # Initial file has freeze_panes=None; golden file has freeze_panes=B1
    try:
        freeze = ws.freeze_panes
        # B1 means col A is frozen; accept also C1 or similar (but B1 is most specific)
        if freeze is not None and str(freeze).startswith('B'):
            print(f"PASS: Component 4 — Freeze panes set to '{freeze}' (column A frozen) (0.10 pts)")
            total_score += 0.10
        elif freeze is not None:
            print(f"FAIL: Component 4 — Freeze panes set to '{freeze}', expected 'B1' to freeze column A")
        else:
            print(f"FAIL: Component 4 — No freeze panes set (expected B1 to freeze column A)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Column A width >= 25 (approximately 200px) (0.05 points)
    # Task requires: "Column A should be at least 200px wide to show full task names"
    # Initial file has col A width=20; golden file has col A width=28
    try:
        col_a_width = ws.column_dimensions.get('A', None)
        width_val = col_a_width.width if col_a_width else None
        # openpyxl uses character units; ~7-8px per char unit. 200px / 7px ≈ 28 chars
        # But we accept anything >= 25 chars as "at least 200px wide"
        if width_val is not None and width_val >= 25:
            print(f"PASS: Component 5 — Column A width={width_val} (>= 25 char units ≈ 200px) (0.05 pts)")
            total_score += 0.05
        elif width_val is not None:
            print(f"FAIL: Component 5 — Column A width={width_val}, expected >= 25 (for ~200px)")
        else:
            print(f"FAIL: Component 5 — Column A width not set")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
