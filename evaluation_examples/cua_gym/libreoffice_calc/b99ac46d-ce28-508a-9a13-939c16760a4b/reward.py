"""
Reward Script: Weekly Flash Report Template Setup
Task ID: calc_fin_weekly_flash_report_080
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: FlashReport header (A1:H1 merged, title text, bold+size16, dark blue bg, white font) — 0.20 pts
  Component 2: FlashReport row 2 week selector (A2 label, B2 TODAY formula) — 0.10 pts
  Component 3: FlashReport row 4 column headers (Metric, Mon..Fri, WTD Total, MTD Total) — 0.10 pts
  Component 4: FlashReport metric labels in rows 5-8 (Revenue, Orders, COGS, Gross Profit) — 0.10 pts
  Component 5: FlashReport VLOOKUP formulas for daily data (B5:F8) — 0.15 pts
  Component 6: FlashReport WTD/MTD aggregation formulas (G5:G8 SUM, H5:H8 SUMPRODUCT) — 0.10 pts
  Component 7: DailyLog Gross Profit formulas (E2:E32 = B-D) — 0.10 pts
  Component 8: Sheet protection on FlashReport (enabled, B2 unlocked) — 0.10 pts
  Component 9: Alternating row shading and number formats on FlashReport — 0.05 pts
  Total: 1.0
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_weekly_flash_report_080'


def verify_task(file_path):
    """
    Verify the weekly flash report template was correctly set up.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify required sheets exist as a precondition gate
    if 'FlashReport' not in wb.sheetnames or 'DailyLog' not in wb.sheetnames:
        print(f"CRITICAL: Required sheets not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws_fr = wb['FlashReport']
    ws_dl = wb['DailyLog']

    # Component 1: FlashReport header row (A1:H1 merged, "Weekly Flash Report", bold, size 16,
    # dark blue background, white font) (0.20 points)
    try:
        # Check title text
        a1_value = ws_fr['A1'].value
        has_title = a1_value is not None and 'weekly flash report' in str(a1_value).lower()

        # Check merge A1:H1
        merged_ranges = [str(mr) for mr in ws_fr.merged_cells.ranges]
        is_merged = 'A1:H1' in merged_ranges

        # Check bold and font size
        a1_cell = ws_fr['A1']
        is_bold = a1_cell.font.bold is True
        has_size_16 = a1_cell.font.size is not None and float(a1_cell.font.size) >= 14

        # Check dark blue background (looking for dark blue-ish color)
        try:
            bg_color = a1_cell.fill.fgColor.rgb
            # Dark blue: FF17375E or similar dark blue ARGB
            # Accept any dark blue variant — R and G are low, B is higher, and it's opaque (FF prefix)
            r_val = int(bg_color[2:4], 16)
            g_val = int(bg_color[4:6], 16)
            b_val = int(bg_color[6:8], 16)
            alpha = bg_color[:2]
            has_dark_blue = alpha == 'FF' and r_val < 60 and g_val < 80 and b_val > 60
        except Exception:
            has_dark_blue = False

        # Check white font color
        try:
            font_color = a1_cell.font.color.rgb
            has_white_font = font_color.upper() in ('FFFFFFFF', 'FFFFFF')
        except Exception:
            has_white_font = False

        if has_title and is_merged and is_bold and has_size_16 and has_dark_blue and has_white_font:
            print(f"PASS: Component 1 — Header merged A1:H1, '{a1_value}', bold size {a1_cell.font.size}, dark blue bg {bg_color}, white font (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 — Header check: title={has_title}({repr(a1_value)}), merged={is_merged}, bold={is_bold}, size16={has_size_16}({a1_cell.font.size}), dark_blue={has_dark_blue}, white_font={has_white_font}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: FlashReport row 2 — "Week of:" label and TODAY()-WEEKDAY formula in B2 (0.10 points)
    try:
        a2_value = ws_fr['A2'].value
        b2_value = ws_fr['B2'].value

        has_week_of = a2_value is not None and 'week of' in str(a2_value).lower()
        # B2 should contain a formula using TODAY() and WEEKDAY()
        has_week_formula = (
            b2_value is not None and
            isinstance(b2_value, str) and
            'TODAY()' in b2_value.upper() and
            'WEEKDAY' in b2_value.upper()
        )

        if has_week_of and has_week_formula:
            print(f"PASS: Component 2 — A2='{a2_value}', B2='{b2_value}' (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 — A2={repr(a2_value)} (week_of={has_week_of}), B2={repr(b2_value)} (week_formula={has_week_formula})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: FlashReport row 4 column headers (0.10 points)
    try:
        expected_headers = ['Metric', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'WTD Total', 'MTD Total']
        actual_headers = [ws_fr.cell(row=4, column=c).value for c in range(1, 9)]

        # Check all expected headers are present (case-insensitive match)
        headers_match = all(
            actual is not None and str(actual).lower() == exp.lower()
            for actual, exp in zip(actual_headers, expected_headers)
        )

        if headers_match:
            print(f"PASS: Component 3 — Row 4 headers correct: {actual_headers} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — Expected {expected_headers}, got {actual_headers}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: FlashReport metric labels in rows 5-8 (0.10 points)
    try:
        expected_metrics = {5: 'Revenue', 6: 'Orders', 7: 'COGS', 8: 'Gross Profit'}
        metrics_ok = True
        actual_metrics = {}

        for row, exp_label in expected_metrics.items():
            actual = ws_fr.cell(row=row, column=1).value
            actual_metrics[row] = actual
            if actual is None or str(actual).strip().lower() != exp_label.lower():
                metrics_ok = False

        if metrics_ok:
            print(f"PASS: Component 4 — Metric labels correct: {list(actual_metrics.values())} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 — Expected {list(expected_metrics.values())}, got {list(actual_metrics.values())}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: FlashReport VLOOKUP formulas for daily data (B5:F8) (0.15 points)
    try:
        vlookup_count = 0
        vlookup_total = 20  # 5 days x 4 metrics

        for row in range(5, 9):
            for col in range(2, 7):  # B to F
                cell_val = ws_fr.cell(row=row, column=col).value
                if (cell_val is not None and
                        isinstance(cell_val, str) and
                        'VLOOKUP' in cell_val.upper() and
                        'DailyLog' in cell_val and
                        'IFERROR' in cell_val.upper()):
                    vlookup_count += 1

        if vlookup_count == vlookup_total:
            print(f"PASS: Component 5 — All {vlookup_total} VLOOKUP/IFERROR formulas present in B5:F8 (0.15 pts)")
            total_score += 0.15
        elif vlookup_count >= vlookup_total // 2:
            partial = 0.07
            print(f"PARTIAL: Component 5 — {vlookup_count}/{vlookup_total} VLOOKUP formulas found (+{partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — Only {vlookup_count}/{vlookup_total} VLOOKUP/IFERROR formulas in B5:F8")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: FlashReport WTD (G5:G8) and MTD (H5:H8) aggregation formulas (0.10 points)
    try:
        wtd_ok = True
        mtd_ok = True

        for row in range(5, 9):
            g_val = ws_fr.cell(row=row, column=7).value  # G column (WTD)
            h_val = ws_fr.cell(row=row, column=8).value  # H column (MTD)

            # G should be SUM(B:F)
            if not (g_val and isinstance(g_val, str) and 'SUM' in g_val.upper() and 'B' in g_val and 'F' in g_val):
                wtd_ok = False

            # H should be SUMPRODUCT or SUMIF referencing DailyLog with month condition
            if not (h_val and isinstance(h_val, str) and
                    ('SUMPRODUCT' in h_val.upper() or 'SUMIF' in h_val.upper()) and
                    'DailyLog' in h_val):
                mtd_ok = False

        if wtd_ok and mtd_ok:
            print(f"PASS: Component 6 — G5:G8 SUM formulas and H5:H8 MTD formulas (SUMPRODUCT/SUMIF) present (0.10 pts)")
            total_score += 0.10
        elif wtd_ok:
            print(f"PARTIAL: Component 6 — G5:G8 SUM OK but H5:H8 MTD formulas missing/incorrect (+0.05 pts)")
            total_score += 0.05
        elif mtd_ok:
            print(f"PARTIAL: Component 6 — H5:H8 MTD OK but G5:G8 SUM formulas missing/incorrect (+0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 — WTD SUM formulas and MTD aggregation formulas not found in G/H columns")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: DailyLog Gross Profit formulas (E2:E32 = B-D formula) (0.10 points)
    try:
        gp_formula_count = 0
        gp_expected = 31  # rows 2-32

        for row in range(2, 33):
            e_val = ws_dl.cell(row=row, column=5).value
            if e_val and isinstance(e_val, str):
                # Should be =B<row>-D<row> style formula
                e_upper = e_val.upper()
                if 'B' in e_upper and 'D' in e_upper and '-' in e_upper:
                    gp_formula_count += 1

        if gp_formula_count >= gp_expected:
            print(f"PASS: Component 7 — DailyLog E2:E32 has {gp_formula_count} Gross Profit formulas (B-D) (0.10 pts)")
            total_score += 0.10
        elif gp_formula_count >= gp_expected // 2:
            partial = 0.05
            print(f"PARTIAL: Component 7 — Only {gp_formula_count}/{gp_expected} Gross Profit formulas in DailyLog E column (+{partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 7 — Only {gp_formula_count}/{gp_expected} Gross Profit formulas found in DailyLog column E")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    # Component 8: Sheet protection on FlashReport (enabled, B2 unlocked for week selector) (0.10 points)
    try:
        fr_protected = ws_fr.protection.sheet is True
        # B2 should be unlocked (locked=False means editable)
        b2_cell = ws_fr['B2']
        b2_unlocked = b2_cell.protection.locked is False

        if fr_protected and b2_unlocked:
            print(f"PASS: Component 8 — FlashReport sheet protection enabled, B2 unlocked (editable) (0.10 pts)")
            total_score += 0.10
        elif fr_protected:
            print(f"PARTIAL: Component 8 — FlashReport protected but B2 is not unlocked (locked={b2_cell.protection.locked}) (+0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 8 — FlashReport protection not enabled (sheet={ws_fr.protection.sheet})")
    except Exception as e:
        print(f"ERROR: Component 8 — {e}")

    # Component 9: Alternating row shading and number formats on FlashReport (0.05 points)
    try:
        # Check alternating fills on rows 5-8
        fills = []
        for row in range(5, 9):
            cell = ws_fr.cell(row=row, column=1)
            try:
                fills.append(cell.fill.fgColor.rgb)
            except Exception:
                fills.append(None)

        # Alternating means row 5 and 7 differ from row 6 and 8
        has_alternating = (
            len(fills) == 4 and
            fills[0] == fills[2] and  # rows 5 and 7 same
            fills[1] == fills[3] and  # rows 6 and 8 same
            fills[0] != fills[1]      # rows 5 and 6 differ
        )

        # Check currency format for Revenue (row 5)
        g5_fmt = ws_fr.cell(row=5, column=7).number_format
        has_currency_revenue = '$' in str(g5_fmt) or 'currency' in str(g5_fmt).lower()

        # Check number format for Orders (row 6) — should be integer, not currency
        g6_fmt = ws_fr.cell(row=6, column=7).number_format
        has_number_orders = '$' not in str(g6_fmt)

        if has_alternating and has_currency_revenue and has_number_orders:
            print(f"PASS: Component 9 — Alternating shading ({fills[0]} vs {fills[1]}), currency/number formats correct (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 9 — alternating={has_alternating}(fills={fills}), currency_revenue={has_currency_revenue}({g5_fmt}), number_orders={has_number_orders}({g6_fmt})")
    except Exception as e:
        print(f"ERROR: Component 9 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
