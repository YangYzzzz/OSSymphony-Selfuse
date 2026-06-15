"""
Initial Setup: Weekly Flash Report Template
Task ID: calc_fin_weekly_flash_report_080
Domain: libreoffice_calc

Creates the pre-task state:
- Sheet 'DailyLog' with headers and dates pre-filled for current month;
  Revenue, Orders, COGS columns empty; Gross Profit column empty (no formula)
- Sheet 'FlashReport' completely blank
- No sheet protection
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date, timedelta
import calendar

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_weekly_flash_report_080'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def get_month_dates():
    """Return list of all dates in current month."""
    today = date.today()
    year = today.year
    month = today.month
    days_in_month = calendar.monthrange(year, month)[1]
    return [date(year, month, d) for d in range(1, days_in_month + 1)]


def create_initial():
    wb = openpyxl.Workbook()

    # -------------------------------------------------------
    # Sheet 1: DailyLog
    # -------------------------------------------------------
    ws_log = wb.active
    ws_log.title = 'DailyLog'

    # Headers in row 1
    headers = ['Date', 'Revenue', 'Orders', 'COGS', 'Gross Profit']
    header_font = Font(name='Calibri', bold=True, size=11)
    header_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    header_border_bottom = Border(
        bottom=Side(style='medium', color='FF000000')
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws_log.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = header_border_bottom

    # Set column widths
    ws_log.column_dimensions['A'].width = 14  # Date
    ws_log.column_dimensions['B'].width = 14  # Revenue
    ws_log.column_dimensions['C'].width = 10  # Orders
    ws_log.column_dimensions['D'].width = 14  # COGS
    ws_log.column_dimensions['E'].width = 14  # Gross Profit

    # Row 1 height
    ws_log.row_dimensions[1].height = 20

    # Pre-fill dates for current month (Rows 2-32)
    month_dates = get_month_dates()
    # Always fill rows 2-32 (pad with dates or leave blank if month < 31 days)
    for row_idx in range(2, 33):
        day_offset = row_idx - 2
        if day_offset < len(month_dates):
            d = month_dates[day_offset]
            date_cell = ws_log.cell(row=row_idx, column=1, value=d)
            date_cell.number_format = 'yyyy-mm-dd'
            date_cell.font = Font(name='Calibri', size=11)
            date_cell.alignment = Alignment(horizontal='center')
        # Columns B (Revenue), C (Orders), D (COGS) — intentionally empty
        # Column E (Gross Profit) — intentionally empty (no formula yet)

    # Freeze header row
    ws_log.freeze_panes = 'A2'

    # -------------------------------------------------------
    # Sheet 2: FlashReport (completely blank)
    # -------------------------------------------------------
    ws_flash = wb.create_sheet('FlashReport')
    # Intentionally blank — no content, no formatting, no protection

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('  Sheet DailyLog: headers + dates rows 2-32 (B/C/D/E empty)')
    print('  Sheet FlashReport: blank')
    print('  No sheet protection')


create_initial()
