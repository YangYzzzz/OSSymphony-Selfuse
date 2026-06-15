"""
Reward Script: Analyze freight costs by carrier and shipping mode.
Task ID: calc_ops_logistics_freight_cost_046
Domain: libreoffice_calc
Scoring:
  Component 1: F2:F101 cost-per-kg formulas (=E/D) — 0.25 pts
  Component 2: G2:G101 contracted-rate lookup formulas (INDEX/MATCH or VLOOKUP) — 0.25 pts
  Component 3: H2:H101 billing-variance formulas (=(F-G)/G) — 0.20 pts
  Component 4: I2:I101 overbill-flag formulas (IF H>2% then OVERBILLED else OK) — 0.15 pts
  Component 5: Red fill (FFFF0000) on OVERBILLED rows — 0.10 pts
  Component 6: K2 summary SUMPRODUCT formula for total overbilling amount — 0.05 pts
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_logistics_freight_cost_046'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify FreightInvoices sheet exists (precondition gate)
    if 'FreightInvoices' not in wb.sheetnames:
        print("CRITICAL: 'FreightInvoices' sheet not found. Cannot score.")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['FreightInvoices']

    # Component 1: F2:F101 cost-per-kg formulas (=E{row}/D{row}) — 0.25 pts
    # Each row should have =E<N>/D<N> formula in column F
    try:
        f_formula_count = 0
        f_total = 100
        for row in range(2, 102):
            val = ws.cell(row=row, column=6).value
            expected = f'=E{row}/D{row}'
            if val and isinstance(val, str) and val.strip().upper() == expected.upper():
                f_formula_count += 1
        coverage = f_formula_count / f_total
        if coverage >= 0.95:
            print(f"PASS: Component 1 — cost-per-kg formulas: {f_formula_count}/100 rows have =E/D formula (0.25 pts)")
            total_score += 0.25
        elif coverage >= 0.5:
            partial = round(0.25 * coverage, 3)
            print(f"PARTIAL: Component 1 — cost-per-kg formulas: {f_formula_count}/100 rows correct ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — only {f_formula_count}/100 rows have =E/D formula in column F")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: G2:G101 contracted-rate lookup formulas — 0.25 pts
    # Should use INDEX/MATCH or VLOOKUP to look up rate from RateCard by Carrier+ShipMode
    try:
        g_formula_count = 0
        g_total = 100
        for row in range(2, 102):
            val = ws.cell(row=row, column=7).value
            if val and isinstance(val, str):
                val_upper = val.strip().upper()
                # Accept INDEX/MATCH or VLOOKUP or any formula referencing RateCard
                if 'RATECARD' in val_upper and (
                    'INDEX' in val_upper or 'VLOOKUP' in val_upper or 'MATCH' in val_upper
                ):
                    g_formula_count += 1
        coverage = g_formula_count / g_total
        if coverage >= 0.95:
            print(f"PASS: Component 2 — contracted-rate lookup formulas: {g_formula_count}/100 rows reference RateCard (0.25 pts)")
            total_score += 0.25
        elif coverage >= 0.5:
            partial = round(0.25 * coverage, 3)
            print(f"PARTIAL: Component 2 — contracted-rate lookup: {g_formula_count}/100 rows correct ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — only {g_formula_count}/100 rows have RateCard lookup in column G")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: H2:H101 billing-variance formulas (=(F{row}-G{row})/G{row}) — 0.20 pts
    try:
        h_formula_count = 0
        h_total = 100
        for row in range(2, 102):
            val = ws.cell(row=row, column=8).value
            if val and isinstance(val, str):
                val_norm = val.strip().upper().replace(' ', '')
                expected = f'=(F{row}-G{row})/G{row}'.upper()
                if val_norm == expected:
                    h_formula_count += 1
        coverage = h_formula_count / h_total
        if coverage >= 0.95:
            print(f"PASS: Component 3 — variance formulas: {h_formula_count}/100 rows have =(F-G)/G formula (0.20 pts)")
            total_score += 0.20
        elif coverage >= 0.5:
            partial = round(0.20 * coverage, 3)
            print(f"PARTIAL: Component 3 — variance formulas: {h_formula_count}/100 rows correct ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — only {h_formula_count}/100 rows have =(F-G)/G formula in column H")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: I2:I101 overbill-flag formulas — 0.15 pts
    # Should be =IF(H{row}>0.02,"OVERBILLED","OK")
    try:
        i_formula_count = 0
        i_total = 100
        for row in range(2, 102):
            val = ws.cell(row=row, column=9).value
            if val and isinstance(val, str):
                val_norm = val.strip().upper().replace(' ', '')
                expected = f'=IF(H{row}>0.02,"OVERBILLED","OK")'.upper()
                if val_norm == expected:
                    i_formula_count += 1
                else:
                    # Also accept variant: >2% or >2/100 or 0.02 with OVERBILLED keyword
                    if 'OVERBILLED' in val_norm and 'IF' in val_norm and 'H' + str(row) in val_norm:
                        i_formula_count += 1
        coverage = i_formula_count / i_total
        if coverage >= 0.95:
            print(f"PASS: Component 4 — overbill-flag formulas: {i_formula_count}/100 rows have IF overbill check (0.15 pts)")
            total_score += 0.15
        elif coverage >= 0.5:
            partial = round(0.15 * coverage, 3)
            print(f"PARTIAL: Component 4 — overbill-flag formulas: {i_formula_count}/100 rows correct ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — only {i_formula_count}/100 rows have IF overbill formula in column I")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Red fill (FFFF0000) on OVERBILLED rows — 0.10 pts
    # The golden file has exactly 16 rows with red fill. We check that at least some rows
    # have red fill in column A (or any column), indicating OVERBILLED rows are highlighted.
    try:
        red_fill_rows = []
        for row in range(2, 102):
            fill = ws.cell(row=row, column=1).fill
            try:
                fgColor = fill.fgColor.rgb
                if fgColor == 'FFFF0000':
                    red_fill_rows.append(row)
            except Exception:
                pass
        red_count = len(red_fill_rows)
        if red_count >= 10:
            print(f"PASS: Component 5 — red fill on OVERBILLED rows: {red_count} rows have red fill (0.10 pts)")
            total_score += 0.10
        elif red_count >= 1:
            partial = round(0.10 * (red_count / 16), 3)
            print(f"PARTIAL: Component 5 — only {red_count} rows have red fill ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — no rows have FFFF0000 red fill in column A")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: K2 summary SUMPRODUCT formula for total overbilling — 0.05 pts
    # K1 should be labeled 'Total Overbilling:' and K2 should have a SUMPRODUCT formula
    try:
        k1_label = ws.cell(row=1, column=11).value
        k2_formula = ws.cell(row=2, column=11).value
        label_ok = k1_label is not None and 'OVERBILL' in str(k1_label).upper()
        formula_ok = (
            k2_formula is not None
            and isinstance(k2_formula, str)
            and ('SUMPRODUCT' in k2_formula.upper() or 'SUM' in k2_formula.upper())
            and 'OVERBILLED' in k2_formula.upper()
        )
        if label_ok and formula_ok:
            print(f"PASS: Component 6 — K1 label='{k1_label}', K2 has summary SUMPRODUCT formula (0.05 pts)")
            total_score += 0.05
        elif formula_ok:
            print(f"PASS: Component 6 — K2 has summary formula (label missing but formula present) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 — K1={repr(k1_label)}, K2={repr(k2_formula)}")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.4f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
