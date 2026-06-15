"""
Initial Setup: Freight Cost Analysis Spreadsheet
Task ID: calc_ops_logistics_freight_cost_046
Domain: libreoffice_calc

Creates:
  - Sheet 'FreightInvoices': 100 shipment records (columns A-I, with F-I header only, data empty)
  - Sheet 'RateCard': contracted rates for each carrier x shipping mode combination

Rate card:
  FastFreight Air  = $21.50/kg
  FastFreight Road = $10.80/kg
  FastFreight Sea  = $ 4.85/kg
  GlobalShip  Air  = $22.20/kg
  GlobalShip  Road = $ 9.90/kg
  GlobalShip  Sea  = $ 5.10/kg
  LocalHaul   Air  = $21.20/kg
  LocalHaul   Road = $ 5.80/kg
  LocalHaul   Sea  = $ 4.60/kg

Overbilled rows (>2% variance): 1, 3, 8, 14, 22, 27, 35, 40, 48, 55, 62, 70, 77, 85, 91, 99
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_logistics_freight_cost_046'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    # -----------------------------------------------------------------------
    # Sheet 1: FreightInvoices
    # -----------------------------------------------------------------------
    ws = wb.active
    ws.title = 'FreightInvoices'

    # Column headers
    headers = [
        'Invoice #',               # A
        'Carrier',                 # B
        'Ship Mode',               # C
        'Weight kg',               # D
        'Billed Amount',           # E
        'Cost per kg',             # F  — EMPTY (task asks agent to fill)
        'Contracted Rate per kg',  # G  — EMPTY (task asks agent to fill)
        'Billing Variance',        # H  — EMPTY (task asks agent to fill)
        'Overbill Flag',           # I  — EMPTY (task asks agent to fill)
    ]
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF', name='Calibri', size=11)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Contracted rates used for computing billed amounts
    # (actual billed may be slightly over or under for realistic variation)
    rate_card = {
        ('FastFreight', 'Air'):   21.50,
        ('FastFreight', 'Road'):  10.80,
        ('FastFreight', 'Sea'):    4.85,
        ('GlobalShip',  'Air'):   22.20,
        ('GlobalShip',  'Road'):   9.90,
        ('GlobalShip',  'Sea'):    5.10,
        ('LocalHaul',   'Air'):   21.20,
        ('LocalHaul',   'Road'):   5.80,
        ('LocalHaul',   'Sea'):    4.60,
    }

    # Invoice data: (Carrier, Ship Mode, Weight kg, variance_pct)
    # variance_pct > 2.0 means overbilled; negative means underbilled
    invoice_specs = [
        # Row 2  — OVERBILLED (+3.5%)
        ('FastFreight', 'Air',   320.5,  3.5),
        # Row 3  — OK (+1.0%)
        ('GlobalShip',  'Sea',   850.0,  1.0),
        # Row 4  — OVERBILLED (+4.2%)
        ('LocalHaul',   'Road',  145.0,  4.2),
        # Row 5  — OK (0.0%)
        ('FastFreight', 'Road',  210.0,  0.0),
        # Row 6  — OK (0.0%)
        ('GlobalShip',  'Air',   480.0,  0.0),
        # Row 7  — OK (-1.5%)
        ('LocalHaul',   'Sea',   620.0, -1.5),
        # Row 8  — OK (+0.8%)
        ('FastFreight', 'Sea',   990.0,  0.8),
        # Row 9  — OVERBILLED (+5.0%)
        ('GlobalShip',  'Road',  275.0,  5.0),
        # Row 10 — OK (0.0%)
        ('LocalHaul',   'Air',   165.0,  0.0),
        # Row 11 — OK (+1.5%)
        ('FastFreight', 'Air',   540.0,  1.5),
        # Row 12 — OK (0.0%)
        ('GlobalShip',  'Sea',  1200.0,  0.0),
        # Row 13 — OK (-2.0%)
        ('LocalHaul',   'Road',  330.0, -2.0),
        # Row 14 — OK (+0.5%)
        ('FastFreight', 'Road',  410.0,  0.5),
        # Row 15 — OVERBILLED (+3.1%)
        ('GlobalShip',  'Air',   290.0,  3.1),
        # Row 16 — OK (0.0%)
        ('LocalHaul',   'Sea',   780.0,  0.0),
        # Row 17 — OK (+1.8%)
        ('FastFreight', 'Sea',  1050.0,  1.8),
        # Row 18 — OK (0.0%)
        ('GlobalShip',  'Road',  185.0,  0.0),
        # Row 19 — OK (-0.5%)
        ('LocalHaul',   'Air',   225.0, -0.5),
        # Row 20 — OK (+1.0%)
        ('FastFreight', 'Air',   310.0,  1.0),
        # Row 21 — OK (0.0%)
        ('GlobalShip',  'Sea',   670.0,  0.0),
        # Row 22 — OK (+0.3%)
        ('LocalHaul',   'Road',  490.0,  0.3),
        # Row 23 — OVERBILLED (+6.0%)
        ('FastFreight', 'Road',  155.0,  6.0),
        # Row 24 — OK (0.0%)
        ('GlobalShip',  'Air',   380.0,  0.0),
        # Row 25 — OK (-1.0%)
        ('LocalHaul',   'Sea',   520.0, -1.0),
        # Row 26 — OK (+1.2%)
        ('FastFreight', 'Sea',   740.0,  1.2),
        # Row 27 — OK (0.0%)
        ('GlobalShip',  'Road',  640.0,  0.0),
        # Row 28 — OVERBILLED (+2.5%)
        ('LocalHaul',   'Air',   195.0,  2.5),
        # Row 29 — OK (+0.7%)
        ('FastFreight', 'Air',   430.0,  0.7),
        # Row 30 — OK (0.0%)
        ('GlobalShip',  'Sea',   910.0,  0.0),
        # Row 31 — OK (-0.8%)
        ('LocalHaul',   'Road',  270.0, -0.8),
        # Row 32 — OK (+1.4%)
        ('FastFreight', 'Road',  380.0,  1.4),
        # Row 33 — OK (0.0%)
        ('GlobalShip',  'Air',   510.0,  0.0),
        # Row 34 — OK (+0.6%)
        ('LocalHaul',   'Sea',   430.0,  0.6),
        # Row 35 — OVERBILLED (+4.8%)
        ('FastFreight', 'Sea',   880.0,  4.8),
        # Row 36 — OK (0.0%)
        ('GlobalShip',  'Road',  320.0,  0.0),
        # Row 37 — OK (+1.9%)
        ('LocalHaul',   'Air',   285.0,  1.9),
        # Row 38 — OK (0.0%)
        ('FastFreight', 'Air',   265.0,  0.0),
        # Row 39 — OK (-0.3%)
        ('GlobalShip',  'Sea',   550.0, -0.3),
        # Row 40 — OK (+0.4%)
        ('LocalHaul',   'Road',  415.0,  0.4),
        # Row 41 — OVERBILLED (+3.7%)
        ('FastFreight', 'Road',  295.0,  3.7),
        # Row 42 — OK (0.0%)
        ('GlobalShip',  'Air',   340.0,  0.0),
        # Row 43 — OK (+1.1%)
        ('LocalHaul',   'Sea',   690.0,  1.1),
        # Row 44 — OK (0.0%)
        ('FastFreight', 'Sea',   960.0,  0.0),
        # Row 45 — OK (+0.2%)
        ('GlobalShip',  'Road',  455.0,  0.2),
        # Row 46 — OK (-1.5%)
        ('LocalHaul',   'Air',   175.0, -1.5),
        # Row 47 — OK (+1.7%)
        ('FastFreight', 'Air',   395.0,  1.7),
        # Row 48 — OVERBILLED (+7.3%)
        ('GlobalShip',  'Sea',   730.0,  7.3),
        # Row 49 — OK (0.0%)
        ('LocalHaul',   'Road',  560.0,  0.0),
        # Row 50 — OK (+0.9%)
        ('FastFreight', 'Road',  225.0,  0.9),
        # Row 51 — OK (0.0%)
        ('GlobalShip',  'Air',   420.0,  0.0),
        # Row 52 — OK (-0.7%)
        ('LocalHaul',   'Sea',   840.0, -0.7),
        # Row 53 — OK (+1.3%)
        ('FastFreight', 'Sea',   610.0,  1.3),
        # Row 54 — OK (0.0%)
        ('GlobalShip',  'Road',  390.0,  0.0),
        # Row 55 — OK (+0.1%)
        ('LocalHaul',   'Air',   240.0,  0.1),
        # Row 56 — OVERBILLED (+3.3%)
        ('FastFreight', 'Air',   470.0,  3.3),
        # Row 57 — OK (0.0%)
        ('GlobalShip',  'Sea',   980.0,  0.0),
        # Row 58 — OK (-1.2%)
        ('LocalHaul',   'Road',  305.0, -1.2),
        # Row 59 — OK (+0.8%)
        ('FastFreight', 'Road',  345.0,  0.8),
        # Row 60 — OK (0.0%)
        ('GlobalShip',  'Air',   260.0,  0.0),
        # Row 61 — OK (+1.6%)
        ('LocalHaul',   'Sea',   590.0,  1.6),
        # Row 62 — OK (0.0%)
        ('FastFreight', 'Sea',   820.0,  0.0),
        # Row 63 — OVERBILLED (+4.1%)
        ('GlobalShip',  'Road',  510.0,  4.1),
        # Row 64 — OK (-0.4%)
        ('LocalHaul',   'Air',   355.0, -0.4),
        # Row 65 — OK (+1.0%)
        ('FastFreight', 'Air',   285.0,  1.0),
        # Row 66 — OK (0.0%)
        ('GlobalShip',  'Sea',  1100.0,  0.0),
        # Row 67 — OK (+0.5%)
        ('LocalHaul',   'Road',  440.0,  0.5),
        # Row 68 — OK (-1.8%)
        ('FastFreight', 'Road',  260.0, -1.8),
        # Row 69 — OK (0.0%)
        ('GlobalShip',  'Air',   315.0,  0.0),
        # Row 70 — OK (+1.9%)
        ('LocalHaul',   'Sea',   760.0,  1.9),
        # Row 71 — OVERBILLED (+5.5%)
        ('FastFreight', 'Sea',   690.0,  5.5),
        # Row 72 — OK (0.0%)
        ('GlobalShip',  'Road',  280.0,  0.0),
        # Row 73 — OK (+0.6%)
        ('LocalHaul',   'Air',   310.0,  0.6),
        # Row 74 — OK (-0.9%)
        ('FastFreight', 'Air',   360.0, -0.9),
        # Row 75 — OK (0.0%)
        ('GlobalShip',  'Sea',   640.0,  0.0),
        # Row 76 — OK (+1.4%)
        ('LocalHaul',   'Road',  375.0,  1.4),
        # Row 77 — OVERBILLED (+2.8%)
        ('FastFreight', 'Road',  480.0,  2.8),
        # Row 78 — OK (0.0%)
        ('GlobalShip',  'Air',   570.0,  0.0),
        # Row 79 — OK (-0.6%)
        ('LocalHaul',   'Sea',   495.0, -0.6),
        # Row 80 — OK (+1.1%)
        ('FastFreight', 'Sea',   550.0,  1.1),
        # Row 81 — OK (0.0%)
        ('GlobalShip',  'Road',  430.0,  0.0),
        # Row 82 — OK (+0.3%)
        ('LocalHaul',   'Air',   200.0,  0.3),
        # Row 83 — OK (-1.1%)
        ('FastFreight', 'Air',   245.0, -1.1),
        # Row 84 — OK (0.0%)
        ('GlobalShip',  'Sea',   870.0,  0.0),
        # Row 85 — OK (+1.7%)
        ('LocalHaul',   'Road',  515.0,  1.7),
        # Row 86 — OVERBILLED (+3.9%)
        ('FastFreight', 'Road',  175.0,  3.9),
        # Row 87 — OK (0.0%)
        ('GlobalShip',  'Air',   445.0,  0.0),
        # Row 88 — OK (-0.2%)
        ('LocalHaul',   'Sea',   635.0, -0.2),
        # Row 89 — OK (+0.7%)
        ('FastFreight', 'Sea',   770.0,  0.7),
        # Row 90 — OK (0.0%)
        ('GlobalShip',  'Road',  360.0,  0.0),
        # Row 91 — OK (+1.8%)
        ('LocalHaul',   'Air',   275.0,  1.8),
        # Row 92 — OVERBILLED (+6.2%)
        ('FastFreight', 'Air',   415.0,  6.2),
        # Row 93 — OK (0.0%)
        ('GlobalShip',  'Sea',   750.0,  0.0),
        # Row 94 — OK (+0.4%)
        ('LocalHaul',   'Road',  465.0,  0.4),
        # Row 95 — OK (-1.3%)
        ('FastFreight', 'Road',  320.0, -1.3),
        # Row 96 — OK (0.0%)
        ('GlobalShip',  'Air',   395.0,  0.0),
        # Row 97 — OK (+1.2%)
        ('LocalHaul',   'Sea',   710.0,  1.2),
        # Row 98 — OK (-0.5%)
        ('FastFreight', 'Sea',   840.0, -0.5),
        # Row 99 — OVERBILLED (+4.4%)
        ('GlobalShip',  'Road',  500.0,  4.4),
        # Row 100 — OK (0.0%)
        ('LocalHaul',   'Air',   340.0,  0.0),
        # Row 101 — OK (+1.5%)
        ('FastFreight', 'Air',   505.0,  1.5),
    ]

    for idx, (carrier, mode, weight, var_pct) in enumerate(invoice_specs):
        row_idx = idx + 2
        inv_num = f'INV-2025-{idx+1:03d}'
        contracted = rate_card[(carrier, mode)]
        # Billed = contracted * weight * (1 + var_pct/100), rounded to 2dp
        billed = round(contracted * weight * (1 + var_pct / 100.0), 2)

        ws.cell(row=row_idx, column=1, value=inv_num)
        ws.cell(row=row_idx, column=2, value=carrier)
        ws.cell(row=row_idx, column=3, value=mode)
        ws.cell(row=row_idx, column=4, value=weight)
        ws.cell(row=row_idx, column=5, value=billed)
        # Columns F, G, H, I left EMPTY intentionally

    # Format numeric columns
    for row_idx in range(2, 102):
        ws.cell(row=row_idx, column=4).number_format = '#,##0.0'
        ws.cell(row=row_idx, column=5).number_format = '$#,##0.00'

    # Column widths
    col_widths = [16, 14, 10, 12, 16, 16, 22, 18, 14]
    for col_idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w

    # Freeze header row
    ws.freeze_panes = 'A2'

    # -----------------------------------------------------------------------
    # Sheet 2: RateCard
    # -----------------------------------------------------------------------
    ws2 = wb.create_sheet('RateCard')

    rate_headers = ['Carrier', 'Ship Mode', 'Rate per kg']
    rate_header_fill = PatternFill(start_color='FF70AD47', end_color='FF70AD47', fill_type='solid')
    rate_header_font = Font(bold=True, color='FFFFFFFF', name='Calibri', size=11)
    for col_idx, h in enumerate(rate_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.fill = rate_header_fill
        cell.font = rate_header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    rate_data = [
        ('FastFreight', 'Air',  21.50),
        ('FastFreight', 'Road', 10.80),
        ('FastFreight', 'Sea',   4.85),
        ('GlobalShip',  'Air',  22.20),
        ('GlobalShip',  'Road',  9.90),
        ('GlobalShip',  'Sea',   5.10),
        ('LocalHaul',   'Air',  21.20),
        ('LocalHaul',   'Road',  5.80),
        ('LocalHaul',   'Sea',   4.60),
    ]

    for row_idx, (carrier, mode, rate) in enumerate(rate_data, 2):
        ws2.cell(row=row_idx, column=1, value=carrier)
        ws2.cell(row=row_idx, column=2, value=mode)
        ws2.cell(row=row_idx, column=3, value=rate)
        ws2.cell(row=row_idx, column=3).number_format = '$#,##0.00'

    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 14

    # Save
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('  Sheet FreightInvoices: 100 rows of freight invoices (columns F-I empty)')
    print('  Sheet RateCard: 9 contracted rate entries')
    print('  Overbilled rows (>2%): rows 2,4,9,15,23,28,35,41,48,56,63,71,77,86,92,99')

create_initial()
