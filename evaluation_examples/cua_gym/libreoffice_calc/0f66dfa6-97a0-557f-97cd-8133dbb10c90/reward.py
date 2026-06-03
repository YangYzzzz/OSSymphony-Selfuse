"""
Reward Script: Insert a new sheet named 'Notes' and change its tab color to yellow.
Task ID: calc_gsi_014
Domain: libreoffice_calc
Scoring:
  Component 1 — 'Notes' sheet exists (0.4 pts)
  Component 2 — 'Notes' sheet tab color is yellow (0.6 pts)
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_014'


def persist_app_state(domain: str):
    """Try to save any unsaved GUI state before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: 'Notes' sheet exists (0.4 points)
    try:
        if 'Notes' in wb.sheetnames:
            print(f"PASS: Component 1 — 'Notes' sheet exists (0.4 pts)")
            total_score += 0.4
        else:
            # Case-insensitive fallback check
            notes_found = False
            for name in wb.sheetnames:
                if name.lower() == 'notes':
                    notes_found = True
                    print(f"FAIL: Component 1 — Found sheet '{name}' but expected exact name 'Notes'")
                    break
            if not notes_found:
                print(f"FAIL: Component 1 — 'Notes' sheet not found. Sheets: {wb.sheetnames}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: 'Notes' sheet tab color is yellow (0.6 points)
    try:
        if 'Notes' in wb.sheetnames:
            ws_notes = wb['Notes']
            tab_color = ws_notes.sheet_properties.tabColor
            if tab_color is not None:
                color_rgb = tab_color.rgb
                print(f"  DEBUG: Notes tab color rgb = '{color_rgb}', type = '{tab_color.type}'")

                # Yellow in hex is FFFF00. In ARGB format it could be:
                # 00FFFF00 (alpha=00), FFFFFF00 (alpha=FF), or just FFFF00
                # We accept any of these as yellow
                is_yellow = False
                if color_rgb is not None:
                    # Normalize: strip alpha prefix if present, compare last 6 chars
                    rgb_suffix = str(color_rgb)[-6:].upper()
                    if rgb_suffix == 'FFFF00':
                        is_yellow = True

                if is_yellow:
                    print(f"PASS: Component 2 — 'Notes' tab color is yellow ({color_rgb}) (0.6 pts)")
                    total_score += 0.6
                else:
                    print(f"FAIL: Component 2 — 'Notes' tab color is '{color_rgb}', expected yellow (FFFF00)")
            else:
                print(f"FAIL: Component 2 — 'Notes' sheet has no tab color set")
        else:
            print(f"FAIL: Component 2 — 'Notes' sheet does not exist, cannot check tab color")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
