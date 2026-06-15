"""
Initial Setup: Insert a new sheet named 'Notes' and change its tab color to yellow.
Task ID: calc_gsi_014
Domain: libreoffice_calc

Creates a project tracking workbook with Tasks, Timeline, Resources, and Budget sheets.
Does NOT include a 'Notes' sheet or any yellow tab colors.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_014'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # --- Sheet 1: Tasks ---
    ws_tasks = wb.active
    ws_tasks.title = "Tasks"
    task_headers = ["Task ID", "Task Name", "Assigned To", "Priority", "Status", "Due Date"]
    for col, h in enumerate(task_headers, 1):
        c = ws_tasks.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = thin_border

    task_data = [
        ["T-001", "Requirements Gathering", "Sarah Chen", "High", "Completed", "2025-02-15"],
        ["T-002", "System Architecture Design", "Marcus Johnson", "High", "Completed", "2025-03-01"],
        ["T-003", "Database Schema Design", "Priya Patel", "High", "In Progress", "2025-03-20"],
        ["T-004", "Frontend Wireframes", "Alex Rivera", "Medium", "In Progress", "2025-03-25"],
        ["T-005", "API Development", "Marcus Johnson", "High", "Not Started", "2025-04-10"],
        ["T-006", "User Authentication Module", "Wei Zhang", "High", "Not Started", "2025-04-15"],
        ["T-007", "Payment Integration", "Sarah Chen", "Medium", "Not Started", "2025-04-30"],
        ["T-008", "Testing & QA", "Priya Patel", "High", "Not Started", "2025-05-15"],
        ["T-009", "Documentation", "Alex Rivera", "Low", "Not Started", "2025-05-20"],
        ["T-010", "Deployment & Go-Live", "Wei Zhang", "High", "Not Started", "2025-06-01"],
        ["T-011", "Post-Launch Monitoring", "Marcus Johnson", "Medium", "Not Started", "2025-06-15"],
        ["T-012", "User Training Sessions", "Sarah Chen", "Medium", "Not Started", "2025-06-20"],
    ]
    for r, row_data in enumerate(task_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_tasks.cell(row=r, column=c, value=val)
            cell.border = thin_border

    ws_tasks.column_dimensions["A"].width = 10
    ws_tasks.column_dimensions["B"].width = 30
    ws_tasks.column_dimensions["C"].width = 20
    ws_tasks.column_dimensions["D"].width = 12
    ws_tasks.column_dimensions["E"].width = 15
    ws_tasks.column_dimensions["F"].width = 14

    # --- Sheet 2: Timeline ---
    ws_timeline = wb.create_sheet("Timeline")
    tl_headers = ["Phase", "Start Date", "End Date", "Duration (Days)", "Milestone"]
    for col, h in enumerate(tl_headers, 1):
        c = ws_timeline.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = thin_border

    tl_data = [
        ["Planning", "2025-01-15", "2025-02-28", 45, "Requirements Approved"],
        ["Design", "2025-03-01", "2025-03-31", 31, "Architecture Sign-off"],
        ["Development Sprint 1", "2025-04-01", "2025-04-15", 15, "Core API Ready"],
        ["Development Sprint 2", "2025-04-16", "2025-04-30", 15, "Auth & Payments"],
        ["Development Sprint 3", "2025-05-01", "2025-05-10", 10, "Feature Complete"],
        ["Testing", "2025-05-11", "2025-05-25", 15, "QA Approved"],
        ["UAT", "2025-05-26", "2025-06-05", 11, "Client Sign-off"],
        ["Deployment", "2025-06-06", "2025-06-10", 5, "Go-Live"],
        ["Hypercare", "2025-06-11", "2025-06-30", 20, "Stable Operations"],
    ]
    for r, row_data in enumerate(tl_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_timeline.cell(row=r, column=c, value=val)
            cell.border = thin_border

    ws_timeline.column_dimensions["A"].width = 25
    ws_timeline.column_dimensions["B"].width = 14
    ws_timeline.column_dimensions["C"].width = 14
    ws_timeline.column_dimensions["D"].width = 18
    ws_timeline.column_dimensions["E"].width = 25

    # --- Sheet 3: Resources ---
    ws_resources = wb.create_sheet("Resources")
    res_headers = ["Name", "Role", "Department", "Allocation (%)", "Hourly Rate ($)", "Email"]
    for col, h in enumerate(res_headers, 1):
        c = ws_resources.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = thin_border

    res_data = [
        ["Sarah Chen", "Project Manager", "PMO", 100, 95, "sarah.chen@acmetech.com"],
        ["Marcus Johnson", "Lead Developer", "Engineering", 80, 110, "marcus.j@acmetech.com"],
        ["Priya Patel", "QA Engineer", "Quality", 60, 85, "priya.p@acmetech.com"],
        ["Alex Rivera", "UX Designer", "Design", 50, 90, "alex.r@acmetech.com"],
        ["Wei Zhang", "DevOps Engineer", "Infrastructure", 40, 105, "wei.z@acmetech.com"],
        ["Laura Kim", "Business Analyst", "Strategy", 30, 80, "laura.k@acmetech.com"],
        ["David Okonkwo", "Security Specialist", "InfoSec", 25, 120, "david.o@acmetech.com"],
    ]
    for r, row_data in enumerate(res_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_resources.cell(row=r, column=c, value=val)
            cell.border = thin_border

    ws_resources.column_dimensions["A"].width = 20
    ws_resources.column_dimensions["B"].width = 22
    ws_resources.column_dimensions["C"].width = 18
    ws_resources.column_dimensions["D"].width = 16
    ws_resources.column_dimensions["E"].width = 18
    ws_resources.column_dimensions["F"].width = 28

    # --- Sheet 4: Budget ---
    ws_budget = wb.create_sheet("Budget")
    bud_headers = ["Category", "Estimated ($)", "Actual ($)", "Variance ($)", "Status"]
    for col, h in enumerate(bud_headers, 1):
        c = ws_budget.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = thin_border

    bud_data = [
        ["Personnel Costs", 185000, 172500, 12500, "Under Budget"],
        ["Software Licenses", 24000, 26800, -2800, "Over Budget"],
        ["Cloud Infrastructure", 36000, 33200, 2800, "Under Budget"],
        ["Hardware", 15000, 14750, 250, "On Track"],
        ["Training", 8000, 4500, 3500, "Under Budget"],
        ["Travel & Meetings", 12000, 9800, 2200, "Under Budget"],
        ["Contingency Reserve", 25000, 0, 25000, "Unspent"],
        ["External Consulting", 30000, 28500, 1500, "On Track"],
        ["Testing Tools", 10000, 11200, -1200, "Over Budget"],
        ["Miscellaneous", 5000, 3100, 1900, "Under Budget"],
    ]
    for r, row_data in enumerate(bud_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_budget.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c in (2, 3, 4) and r > 1:
                cell.number_format = '#,##0.00'

    ws_budget.column_dimensions["A"].width = 22
    ws_budget.column_dimensions["B"].width = 16
    ws_budget.column_dimensions["C"].width = 14
    ws_budget.column_dimensions["D"].width = 16
    ws_budget.column_dimensions["E"].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
