"""
Initial Setup: Create QoQ performance spreadsheet with employee ratings data.
Task ID: calc_hr_048
Domain: libreoffice_calc

The spreadsheet has employee names and Q3/Q4 ratings. Column D (Change) is
left empty because the task asks the agent to enter formulas and apply
conditional formatting there.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_048'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'QoQ'

    # --- Headers ---
    headers = ['Employee', 'Q3 Rating', 'Q4 Rating', 'Change']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font_white = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = header_alignment

    # --- Employee Data ---
    # Column D (Change) is intentionally left empty - task asks agent to add formulas
    data = [
        ['Alice',  3.8, 4.2],
        ['Bob',    4.1, 3.9],
        ['Carol',  3.5, 3.5],
        ['Dan',    4.0, 4.5],
        ['Eve',    3.2, 2.8],
    ]

    name_font = Font(name='Calibri', size=11)
    number_font = Font(name='Calibri', size=11)

    for r, row_data in enumerate(data, 2):
        # Employee name
        cell_name = ws.cell(row=r, column=1, value=row_data[0])
        cell_name.font = name_font

        # Q3 Rating
        cell_q3 = ws.cell(row=r, column=2, value=row_data[1])
        cell_q3.font = number_font
        cell_q3.number_format = '0.0'
        cell_q3.alignment = Alignment(horizontal='center')

        # Q4 Rating
        cell_q4 = ws.cell(row=r, column=3, value=row_data[2])
        cell_q4.font = number_font
        cell_q4.number_format = '0.0'
        cell_q4.alignment = Alignment(horizontal='center')

        # Column D left empty (agent task)

    # --- Column widths ---
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12

    # --- Freeze header row ---
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
