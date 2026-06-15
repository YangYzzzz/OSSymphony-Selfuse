"""
Reward Script: Conditional formatting with icon sets for quarterly performance change
Task ID: calc_hr_048
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): D2:D6 contain =Cn-Bn formulas
  Component 2 (0.3): Conditional formatting with iconSet rule on D2:D6
  Component 3 (0.3): IconSet uses 3-arrow variant with correct threshold config
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_048'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that 'QoQ' sheet exists (precondition gate)
    if 'QoQ' not in wb.sheetnames:
        print("CRITICAL: Sheet 'QoQ' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['QoQ']

    # Component 1: D2:D6 contain formulas =Cn-Bn (0.4 points)
    # These should be difference formulas: =C2-B2, =C3-B3, etc.
    try:
        formula_count = 0
        expected_formulas = {
            2: '=C2-B2',
            3: '=C3-B3',
            4: '=C4-B4',
            5: '=C5-B5',
            6: '=C6-B6',
        }
        for row, expected in expected_formulas.items():
            cell_val = ws.cell(row=row, column=4).value
            if cell_val is not None and isinstance(cell_val, str):
                # Normalize: remove spaces, uppercase
                normalized = cell_val.upper().replace(" ", "")
                expected_norm = expected.upper().replace(" ", "")
                if normalized == expected_norm:
                    formula_count += 1
                else:
                    print(f"FAIL: Component 1 — D{row} has '{cell_val}', expected '{expected}'")
            else:
                print(f"FAIL: Component 1 — D{row} value is {repr(cell_val)}, expected formula")

        if formula_count == 5:
            print(f"PASS: Component 1 — All 5 formulas correct in D2:D6 (0.4 pts)")
            total_score += 0.4
        elif formula_count > 0:
            partial = round(0.4 * formula_count / 5, 2)
            print(f"PARTIAL: Component 1 — {formula_count}/5 formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No correct formulas found in D2:D6")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Conditional formatting with iconSet rule exists on D2:D6 (0.3 points)
    try:
        icon_set_found = False
        target_range_match = False
        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            for rule in cf.rules:
                if rule.type == 'iconSet' and rule.iconSet is not None:
                    icon_set_found = True
                    # Check if the range covers D2:D6
                    cf_range_upper = cf_range.upper()
                    if 'D2' in cf_range_upper and 'D6' in cf_range_upper:
                        target_range_match = True

        if icon_set_found and target_range_match:
            print(f"PASS: Component 2 — IconSet conditional formatting found on D2:D6 (0.3 pts)")
            total_score += 0.3
        elif icon_set_found:
            print(f"PARTIAL: Component 2 — IconSet found but not on expected range D2:D6 (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 — No iconSet conditional formatting found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: IconSet is a 3-arrow variant with appropriate thresholds (0.3 points)
    try:
        arrow_set_found = False
        threshold_ok = False
        for cf in ws.conditional_formatting:
            for rule in cf.rules:
                if rule.type == 'iconSet' and rule.iconSet is not None:
                    icon = rule.iconSet
                    icon_name = icon.iconSet if hasattr(icon, 'iconSet') else ''
                    # Accept any 3-arrow variant (3Arrows, 3ArrowsGray, etc.)
                    if '3' in str(icon_name) and 'rrow' in str(icon_name).lower():
                        arrow_set_found = True
                        # Check cfvo thresholds: should have 3 cfvo entries
                        # For up/neutral/down based on positive/zero/negative,
                        # the thresholds should reference 0
                        if hasattr(icon, 'cfvo') and icon.cfvo is not None:
                            cfvo_list = list(icon.cfvo)
                            if len(cfvo_list) == 3:
                                # Check that thresholds involve 0 values for
                                # differentiating positive/zero/negative
                                vals = [c.val for c in cfvo_list]
                                if all(v is not None and float(v) == 0.0 for v in vals):
                                    threshold_ok = True

        if arrow_set_found and threshold_ok:
            print(f"PASS: Component 3 — 3Arrows iconSet with correct thresholds (0.3 pts)")
            total_score += 0.3
        elif arrow_set_found:
            print(f"PARTIAL: Component 3 — 3Arrows found but thresholds may not be optimal (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 — No 3-arrow iconSet variant found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
