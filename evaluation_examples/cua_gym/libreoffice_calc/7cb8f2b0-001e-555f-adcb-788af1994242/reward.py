"""
Reward Script: Create a pivot table analyzing quarterly sales data
Task ID: calc_adv_pivot_multifield_003
Domain: libreoffice_calc
Scoring:
  Component 1: Pivot_Table sheet exists (0.3 pts)
  Component 2: Correct headers — Sales Rep rows, Q1/Q2/Q3/Q4 column headers, Grand Total (0.3 pts)
  Component 3: Correct revenue values in pivot table (spot checks + grand total row) (0.4 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_adv_pivot_multifield_003'

# Expected values from the golden file
EXPECTED_SALES_REPS = ['Alice', 'Bob', 'Carol', 'David', 'Eve']
EXPECTED_QUARTERS = ['Q1', 'Q2', 'Q3', 'Q4']
EXPECTED_GRAND_TOTAL_ROW = 1179101.85 + 1218967.89 + 1703828.59 + 1307576.0  # sum of Q totals = 5409474.33
EXPECTED_TOTAL = 5409474.33
# Row-level spot checks (Sales Rep -> {quarter -> revenue, Grand Total})
EXPECTED_ALICE = {'Q1': 129314.88, 'Q2': 279268.28, 'Q3': 124577.12, 'Q4': 223531.59, 'Grand Total': 756691.87}
EXPECTED_GRAND_TOTALS = {'Q1': 1179101.85, 'Q2': 1218967.89, 'Q3': 1703828.59, 'Q4': 1307576.0, 'Grand Total': 5409474.33}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Pivot_Table sheet exists (0.3 points)
    # This FAILS on the initial file (no Pivot_Table sheet) and PASSES on the golden file.
    try:
        if 'Pivot_Table' in wb.sheetnames:
            print("PASS: Component 1 — 'Pivot_Table' sheet exists (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — 'Pivot_Table' sheet not found. Sheets present: {wb.sheetnames}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # If no Pivot_Table sheet, remaining checks cannot proceed
    if 'Pivot_Table' not in wb.sheetnames:
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    ws = wb['Pivot_Table']

    # Component 2: Correct header structure (0.3 points)
    # Check row 3 has: A3='Sales Rep', B3='Q1', C3='Q2', D3='Q3', E3='Q4', F3='Grand Total'
    # Also checks that Sales Rep names appear as row labels in column A (rows 4-8)
    # This FAILS on initial (no sheet) and PASSES on golden.
    try:
        header_checks = 0
        # Check column headers in row 3
        expected_headers = {1: 'Sales Rep', 2: 'Q1', 3: 'Q2', 4: 'Q3', 5: 'Q4', 6: 'Grand Total'}
        header_match = True
        for col, expected in expected_headers.items():
            actual = ws.cell(row=3, column=col).value
            if actual is None or str(actual).strip() != expected:
                print(f"FAIL: Component 2 header check — col {col} expected '{expected}', got '{actual}'")
                header_match = False
                break

        if header_match:
            header_checks += 1

        # Check that Sales Rep row labels in col A (rows 4-8) contain all 5 expected reps
        actual_reps = []
        for row in range(4, 9):
            val = ws.cell(row=row, column=1).value
            if val is not None:
                actual_reps.append(str(val).strip())

        missing_reps = [rep for rep in EXPECTED_SALES_REPS if rep not in actual_reps]
        if not missing_reps:
            header_checks += 1
        else:
            print(f"FAIL: Component 2 row labels — missing sales reps: {missing_reps}. Found: {actual_reps}")

        if header_checks == 2:
            print(f"PASS: Component 2 — Correct headers (Sales Rep, Q1-Q4, Grand Total) and all 5 sales reps as row labels (0.3 pts)")
            total_score += 0.3
        elif header_checks == 1:
            print(f"PARTIAL: Component 2 — Only {header_checks}/2 header sub-checks passed (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 — Header structure incorrect")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Correct revenue values in the pivot table (0.4 points)
    # Verify specific values: Alice's row, Grand Total row.
    # These only differ from the initial file (where no pivot table exists).
    try:
        value_checks = 0
        total_value_checks = 0

        # Build a lookup: row label -> {col header -> row index}
        # Find the row for each sales rep and the column for each quarter
        col_map = {}  # header string -> col index
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=3, column=col).value
            if val is not None:
                col_map[str(val).strip()] = col

        rep_row_map = {}  # rep name -> row index
        for row in range(4, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val is not None:
                rep_row_map[str(val).strip()] = row

        tolerance = 1.0  # allow $1 tolerance for floating point

        # Check Alice's row values
        if 'Alice' in rep_row_map:
            alice_row = rep_row_map['Alice']
            alice_passed = 0
            alice_total = len(EXPECTED_ALICE)
            for quarter, expected_val in EXPECTED_ALICE.items():
                if quarter in col_map:
                    actual = ws.cell(row=alice_row, column=col_map[quarter]).value
                    if actual is not None and abs(float(actual) - expected_val) <= tolerance:
                        alice_passed += 1
                    else:
                        print(f"  Alice {quarter}: expected {expected_val}, got {actual}")

            if alice_passed == alice_total:
                print(f"PASS: Component 3a — Alice's revenue values all correct")
                value_checks += 1
            else:
                print(f"FAIL: Component 3a — Alice's revenue values: {alice_passed}/{alice_total} correct")
            total_value_checks += 1

        # Check Grand Total row values
        if 'Grand Total' in rep_row_map:
            grand_row = rep_row_map['Grand Total']
            grand_passed = 0
            grand_total_checks = len(EXPECTED_GRAND_TOTALS)
            for quarter, expected_val in EXPECTED_GRAND_TOTALS.items():
                if quarter in col_map:
                    actual = ws.cell(row=grand_row, column=col_map[quarter]).value
                    if actual is not None and abs(float(actual) - expected_val) <= tolerance:
                        grand_passed += 1
                    else:
                        print(f"  Grand Total {quarter}: expected {expected_val}, got {actual}")

            if grand_passed == grand_total_checks:
                print(f"PASS: Component 3b — Grand Total row values all correct")
                value_checks += 1
            else:
                print(f"FAIL: Component 3b — Grand Total row values: {grand_passed}/{grand_total_checks} correct")
            total_value_checks += 1

        # Assign score proportionally
        if total_value_checks > 0:
            score_fraction = value_checks / total_value_checks
            component3_score = round(0.4 * score_fraction, 4)
            if score_fraction == 1.0:
                print(f"PASS: Component 3 — All revenue values correct (0.4 pts)")
            else:
                print(f"PARTIAL: Component 3 — {value_checks}/{total_value_checks} value checks passed ({component3_score} pts)")
            total_score += component3_score
        else:
            print(f"FAIL: Component 3 — Could not locate data rows (no 'Alice' or 'Grand Total' rows found)")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
