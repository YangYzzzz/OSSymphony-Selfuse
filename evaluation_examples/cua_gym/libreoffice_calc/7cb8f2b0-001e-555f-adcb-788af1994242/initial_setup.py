"""
Initial Setup: Quarterly Sales Data for Pivot Table Task
Task ID: calc_adv_pivot_multifield_003
Domain: libreoffice_calc

Creates a Q_Sales sheet with 200 rows of transaction data.
No pivot table exists yet — the agent must create one.
"""

import os
import random
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_adv_pivot_multifield_003'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

random.seed(42)

def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Q_Sales ---
    ws = wb.active
    ws.title = 'Q_Sales'

    # Headers: Transaction ID, Sales Rep, Quarter, Product Line, Revenue, Units
    headers = ['Transaction ID', 'Sales Rep', 'Quarter', 'Product Line', 'Revenue', 'Units']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    sales_reps = ['Alice', 'Bob', 'Carol', 'David', 'Eve']
    quarters = ['Q1', 'Q2', 'Q3', 'Q4']
    product_lines = [
        'Enterprise Software', 'Cloud Services', 'Hardware Solutions',
        'Professional Services', 'Support & Maintenance', 'Training Programs'
    ]

    # Revenue ranges per product line (realistic business figures)
    revenue_ranges = {
        'Enterprise Software':    (18000, 95000),
        'Cloud Services':         (5000,  42000),
        'Hardware Solutions':     (8500,  67000),
        'Professional Services':  (12000, 58000),
        'Support & Maintenance':  (3200,  22000),
        'Training Programs':      (1800,  14000),
    }

    # Unit ranges per product line
    unit_ranges = {
        'Enterprise Software':    (1,  5),
        'Cloud Services':         (3, 20),
        'Hardware Solutions':     (2, 15),
        'Professional Services':  (4, 30),
        'Support & Maintenance':  (6, 40),
        'Training Programs':      (5, 25),
    }

    # Generate 200 rows (rows 2-201)
    for i in range(200):
        row = i + 2
        txn_id = f'TXN-2025-{1000 + i:04d}'
        rep = random.choice(sales_reps)
        quarter = random.choice(quarters)
        product = random.choice(product_lines)
        rev_lo, rev_hi = revenue_ranges[product]
        revenue = round(random.uniform(rev_lo, rev_hi), 2)
        unit_lo, unit_hi = unit_ranges[product]
        units = random.randint(unit_lo, unit_hi)

        ws.cell(row=row, column=1, value=txn_id)
        ws.cell(row=row, column=2, value=rep)
        ws.cell(row=row, column=3, value=quarter)
        ws.cell(row=row, column=4, value=product)
        ws.cell(row=row, column=5, value=revenue)
        ws.cell(row=row, column=6, value=units)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Q_Sales with 200 data rows (rows 2-201)')
    print(f'  Columns: Transaction ID, Sales Rep, Quarter, Product Line, Revenue, Units')
    print(f'  No pivot table — agent must create one.')

create_initial()
