"""
Initial Setup: Reading Level Progress Tracker for Elementary Students
Task ID: calc_edu_reading_level_tracker_046
Domain: libreoffice_calc

Creates a spreadsheet with 25 students' reading levels across 3 assessment periods.
Columns E (Growth) and F (Regressed) are empty.
Row 28 has Class Average label but no formulas.
No conditional formatting or charts.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_reading_level_tracker_046'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: ReadingLevels ---
    ws = wb.active
    ws.title = 'ReadingLevels'

    # Headers in row 1
    headers = ['Student', 'Period1 Level', 'Period2 Level', 'Period3 Level', 'Growth', 'Regressed']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # 25 students with realistic names and reading levels (scale 1-30)
    # Reading levels span a realistic range for elementary students
    students_data = [
        ('Amelia Rodriguez',    8,  10, 12),
        ('Benjamin Carter',     6,   7,  9),
        ('Chloe Nguyen',       11,  13, 15),
        ('Daniel Kim',          5,   6,  5),   # regression
        ('Emma Thompson',      14,  16, 18),
        ('Finn O\'Brien',       9,  11, 13),
        ('Grace Patel',        12,  12, 14),
        ('Henry Wallace',       7,   8, 10),
        ('Isabel Morales',     10,  11, 12),
        ('James Lawson',        4,   5,  3),   # regression
        ('Katie Hernandez',    16,  18, 20),
        ('Liam Foster',         8,   9, 11),
        ('Maya Robinson',      13,  15, 17),
        ('Noah Bennett',        6,   7,  8),
        ('Olivia Chen',        18,  20, 22),
        ('Patrick Sullivan',    7,   8,  7),   # regression
        ('Quinn Ramirez',      11,  12, 14),
        ('Riley Anderson',      9,  10, 12),
        ('Samuel Torres',      15,  16, 18),
        ('Tanya Patel',         5,   6,  7),
        ('Uma Krishnan',       12,  14, 15),
        ('Victor Okonkwo',      3,   4,  6),
        ('Wendy Larson',       10,  11, 13),
        ('Xavier Diaz',         8,   9, 10),
        ('Yara Mitchell',      14,  15, 17),
    ]

    for r, (name, p1, p2, p3) in enumerate(students_data, 2):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=p1)
        ws.cell(row=r, column=3, value=p2)
        ws.cell(row=r, column=4, value=p3)
        # Columns E (Growth) and F (Regressed) are left empty - agent must fill these

    # Row 28: Class Average label — no formulas yet (agent must add them)
    ws.cell(row=28, column=1, value='Class Average')
    ws.cell(row=28, column=1).font = Font(bold=True)
    # B28, C28, D28 left empty — agent must add AVERAGE formulas

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: ReadingLevels')
    print(f'  Rows: 1 header + 25 data rows + 1 average row label')
    print(f'  Columns E and F: EMPTY (to be filled by agent)')
    print(f'  Row 28: Class Average label only, no formulas')
    print(f'  No conditional formatting, no chart')


create_initial()
