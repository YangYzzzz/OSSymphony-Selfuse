"""
Reward Script: Track reading level progress for elementary students
Task ID: calc_edu_reading_level_tracker_046
Domain: libreoffice_calc
Scoring:
  Component 1: Growth formulas in column E (rows 2-26)            — 0.25 points
  Component 2: Regressed IF formulas in column F (rows 2-26)      — 0.25 points
  Component 3: AVERAGE formulas in row 28 (B28, C28, D28)         — 0.20 points
  Component 4: Conditional formatting for regressed rows          — 0.15 points
  Component 5: Line chart with correct title                      — 0.15 points
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_edu_reading_level_tracker_046'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'ReadingLevels' not in wb.sheetnames:
        print("CRITICAL: Sheet 'ReadingLevels' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['ReadingLevels']

    # Component 1: Growth formulas in column E (0.25 points)
    # Each student row 2-26 should have =D{row}-B{row}
    # This FAILS on initial (column E is empty) and PASSES on golden
    try:
        growth_correct = 0
        growth_total = 25  # rows 2-26

        for row in range(2, 27):
            val = ws.cell(row=row, column=5).value  # column E
            expected_formula = f'=D{row}-B{row}'
            if val is not None and str(val).strip().upper() == expected_formula.upper():
                growth_correct += 1

        if growth_correct == growth_total:
            print(f"PASS: Component 1 — All {growth_total} Growth formulas correct (=D{{row}}-B{{row}}) (0.25 pts)")
            total_score += 0.25
        elif growth_correct >= growth_total * 0.8:
            print(f"PARTIAL: Component 1 — {growth_correct}/{growth_total} Growth formulas correct (0.12 pts)")
            total_score += 0.12
        else:
            print(f"FAIL: Component 1 — Only {growth_correct}/{growth_total} Growth formulas correct")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Regressed IF formulas in column F (0.25 points)
    # Each student row 2-26 should have =IF(D{row}<B{row},"Yes","No")
    # This FAILS on initial (column F is empty) and PASSES on golden
    try:
        regressed_correct = 0
        regressed_total = 25  # rows 2-26

        for row in range(2, 27):
            val = ws.cell(row=row, column=6).value  # column F
            if val is None:
                continue
            val_str = str(val).strip().upper().replace(' ', '')
            # Accept variations in quoting or case
            expected_base = f'=IF(D{row}<B{row},"YES","NO")'
            if val_str == expected_base:
                regressed_correct += 1
            else:
                # Also accept with different yes/no casing
                import re
                pattern = rf'^=IF\(D{row}<B{row},"YES","NO"\)$'
                if re.match(pattern, val_str, re.IGNORECASE):
                    regressed_correct += 1

        if regressed_correct == regressed_total:
            print(f"PASS: Component 2 — All {regressed_total} Regressed IF formulas correct (0.25 pts)")
            total_score += 0.25
        elif regressed_correct >= regressed_total * 0.8:
            print(f"PARTIAL: Component 2 — {regressed_correct}/{regressed_total} Regressed IF formulas correct (0.12 pts)")
            total_score += 0.12
        else:
            print(f"FAIL: Component 2 — Only {regressed_correct}/{regressed_total} Regressed IF formulas correct")
            if regressed_correct > 0:
                # Show a sample of what was found
                sample = ws.cell(row=2, column=6).value
                print(f"  Sample F2 value: {repr(sample)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: AVERAGE formulas in row 28 for columns B, C, D (0.20 points)
    # B28=AVERAGE(B2:B26), C28=AVERAGE(C2:C26), D28=AVERAGE(D2:D26)
    # This FAILS on initial (B28, C28, D28 are None) and PASSES on golden
    try:
        avg_correct = 0
        expected_averages = {
            'B': '=AVERAGE(B2:B26)',
            'C': '=AVERAGE(C2:C26)',
            'D': '=AVERAGE(D2:D26)',
        }
        col_map = {'B': 2, 'C': 3, 'D': 4}

        for col_letter, expected_formula in expected_averages.items():
            col_idx = col_map[col_letter]
            val = ws.cell(row=28, column=col_idx).value
            if val is not None and str(val).strip().upper() == expected_formula.upper():
                avg_correct += 1
            else:
                print(f"  FAIL: {col_letter}28 expected {expected_formula}, found {repr(val)}")

        if avg_correct == 3:
            print(f"PASS: Component 3 — All 3 AVERAGE formulas in row 28 correct (0.20 pts)")
            total_score += 0.20
        elif avg_correct >= 2:
            print(f"PARTIAL: Component 3 — {avg_correct}/3 AVERAGE formulas correct (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — Only {avg_correct}/3 AVERAGE formulas in row 28 correct")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Conditional formatting (orange fill on regressed rows) (0.15 points)
    # Formula-based CF rule on A2:F26 with formula $F2="Yes" and orange fill
    # This FAILS on initial (no CF rules) and PASSES on golden
    try:
        cf_found = False
        cf_formula_ok = False
        cf_color_ok = False

        cf_rules = ws.conditional_formatting
        for cf in cf_rules:
            for rule in cf.rules:
                if rule.type == 'expression' and hasattr(rule, 'formula') and rule.formula:
                    formula_str = str(rule.formula[0]).strip()
                    # Check formula references column F for "Yes"
                    if 'F' in formula_str.upper() and '"YES"' in formula_str.upper():
                        cf_found = True
                        cf_formula_ok = True
                        # Check fill color is orange-ish
                        if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                            try:
                                fill_rgb = rule.dxf.fill.fgColor.rgb
                                # Orange ARGB: FFFFA500 (or similar orange shades)
                                # Accept any orange-ish color (red high, green medium, blue low)
                                if fill_rgb and len(fill_rgb) == 8:
                                    r_val = int(fill_rgb[2:4], 16)
                                    g_val = int(fill_rgb[4:6], 16)
                                    b_val = int(fill_rgb[6:8], 16)
                                    # Orange: high red, medium-high green, low blue
                                    if r_val >= 200 and g_val >= 100 and b_val <= 100:
                                        cf_color_ok = True
                                        print(f"  CF fill color: {fill_rgb} (orange)")
                                    else:
                                        print(f"  CF fill color: {fill_rgb} (not orange)")
                            except Exception as ce:
                                print(f"  Error reading CF color: {ce}")

        if cf_found and cf_formula_ok and cf_color_ok:
            print(f"PASS: Component 4 — Conditional formatting with orange fill on regressed rows (0.15 pts)")
            total_score += 0.15
        elif cf_found and cf_formula_ok:
            print(f"PARTIAL: Component 4 — CF rule exists with correct formula but fill color issue (0.07 pts)")
            total_score += 0.07
        elif cf_found:
            print(f"PARTIAL: Component 4 — CF rule exists but formula doesn't check column F='Yes' (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4 — No conditional formatting found for regressed rows")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Line chart with correct title (0.15 points)
    # A line chart using average row 28 data, with title 'Class Average Reading Level Progress'
    # This FAILS on initial (no charts) and PASSES on golden
    try:
        charts = ws._charts
        line_chart_found = False
        chart_title_ok = False

        for chart in charts:
            # Check if it's a LineChart
            if type(chart).__name__ in ('LineChart', 'LineChart3D'):
                line_chart_found = True
                # Check chart title
                if chart.title is not None:
                    try:
                        # Extract title text
                        title_text = ''
                        if hasattr(chart.title, 'tx') and chart.title.tx:
                            if hasattr(chart.title.tx, 'rich') and chart.title.tx.rich:
                                for para in chart.title.tx.rich.p:
                                    for run in para.r:
                                        title_text += run.t
                        if not title_text:
                            # Try another approach
                            title_text = str(chart.title)
                        expected_title = 'Class Average Reading Level Progress'
                        if expected_title.lower() in title_text.lower():
                            chart_title_ok = True
                            print(f"  Chart title found: '{title_text}'")
                        else:
                            print(f"  Chart title found but unexpected: '{title_text}'")
                    except Exception as te:
                        print(f"  Error reading chart title: {te}")
                else:
                    print(f"  Line chart found but no title")

        if line_chart_found and chart_title_ok:
            print(f"PASS: Component 5 — Line chart with correct title found (0.15 pts)")
            total_score += 0.15
        elif line_chart_found:
            print(f"PARTIAL: Component 5 — Line chart found but title incorrect (0.07 pts)")
            total_score += 0.07
        else:
            chart_count = len(charts)
            if chart_count > 0:
                chart_types = [type(c).__name__ for c in charts]
                print(f"FAIL: Component 5 — No line chart found (found {chart_count} chart(s): {chart_types})")
            else:
                print(f"FAIL: Component 5 — No charts found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
