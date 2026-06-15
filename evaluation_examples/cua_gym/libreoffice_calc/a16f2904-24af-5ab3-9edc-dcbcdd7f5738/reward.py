"""
Reward Script: Sprint backlog board with story point formatting and burndown tracking
Task ID: calc_gpm_059
Domain: libreoffice_calc
Scoring:
  Component 1: Completion formulas in J4:J15 (0.35 pts)
  Component 2: Sprint Totals row 17 - label, SUM, SUMPRODUCT (0.30 pts)
  Component 3: Burndown line chart present with title and series (0.35 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_059'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Sprint' sheet must exist
    if 'Sprint' not in wb.sheetnames:
        print("FAIL: 'Sprint' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Sprint']

    # =========================================================================
    # Component 1: Completion formulas in J4:J15 (0.35 points)
    # The golden file has IF formulas computing completion % based on Status.
    # Initial file has None in all J4:J15 cells, so this differentiates.
    # =========================================================================
    try:
        formula_count = 0
        correct_formula_count = 0
        for row in range(4, 16):
            cell_val = ws.cell(row=row, column=10).value  # Column J
            if cell_val is not None and isinstance(cell_val, str) and cell_val.startswith('='):
                formula_count += 1
                # Check it references the Status column (E) and uses IF logic
                upper_val = cell_val.upper().replace(' ', '')
                if 'IF(' in upper_val and f'E{row}' in cell_val.upper().replace(' ', ''):
                    correct_formula_count += 1

        if correct_formula_count >= 10:
            # At least 10 of 12 have correct IF formulas referencing E column
            print(f"PASS: Component 1 — {correct_formula_count}/12 completion IF formulas found (0.35 pts)")
            total_score += 0.35
        elif formula_count >= 6:
            # Partial: have formulas but not all correct
            partial = 0.15
            print(f"PARTIAL: Component 1 — {formula_count} formulas found, {correct_formula_count} correct IF formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Expected IF formulas in J4:J15, found {formula_count} formulas, {correct_formula_count} correct")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Sprint Totals row 17 (0.30 points)
    # Golden: A17='Sprint Totals' bold, C17=SUM(C4:C15), D17=SUMPRODUCT
    # Initial: Row 17 is entirely empty
    # Sub-components: label (0.10), SUM formula (0.10), SUMPRODUCT formula (0.10)
    # =========================================================================
    try:
        comp2_score = 0.0

        # Sub 2a: A17 has "Sprint Totals" text and is bold
        a17_val = ws.cell(row=17, column=1).value
        a17_bold = ws.cell(row=17, column=1).font.bold
        if a17_val is not None and 'sprint' in str(a17_val).lower() and 'total' in str(a17_val).lower():
            if a17_bold:
                print(f"PASS: Component 2a — A17 = '{a17_val}', bold=True (0.10 pts)")
                comp2_score += 0.10
            else:
                print(f"PARTIAL: Component 2a — A17 = '{a17_val}' but not bold (0.05 pts)")
                comp2_score += 0.05
        else:
            print(f"FAIL: Component 2a — Expected 'Sprint Totals' in A17, found: {a17_val}")

        # Sub 2b: C17 has SUM formula for story points
        c17_val = ws.cell(row=17, column=3).value
        if c17_val is not None and isinstance(c17_val, str) and '=SUM' in c17_val.upper().replace(' ', ''):
            print(f"PASS: Component 2b — C17 has SUM formula: {c17_val} (0.10 pts)")
            comp2_score += 0.10
        else:
            print(f"FAIL: Component 2b — Expected SUM formula in C17, found: {c17_val}")

        # Sub 2c: D17 has SUMPRODUCT formula for completed points
        d17_val = ws.cell(row=17, column=4).value
        if d17_val is not None and isinstance(d17_val, str) and 'SUMPRODUCT' in d17_val.upper().replace(' ', ''):
            print(f"PASS: Component 2c — D17 has SUMPRODUCT formula: {d17_val} (0.10 pts)")
            comp2_score += 0.10
        else:
            print(f"FAIL: Component 2c — Expected SUMPRODUCT formula in D17, found: {d17_val}")

        total_score += comp2_score
        print(f"  Component 2 total: {comp2_score}/0.30")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Burndown chart (0.35 points)
    # Golden has a LineChart titled 'Sprint Burndown' with 2 series.
    # Initial has 0 charts.
    # Sub-components: chart exists (0.15), is line chart (0.05), has title (0.10), has >=2 series (0.05)
    # =========================================================================
    try:
        charts = ws._charts
        if len(charts) >= 1:
            comp3_score = 0.0
            print(f"PASS: Component 3a — Chart found ({len(charts)} chart(s)) (0.15 pts)")
            comp3_score += 0.15

            chart = charts[0]

            # Check chart type is line
            from openpyxl.chart import LineChart
            if isinstance(chart, LineChart):
                print(f"PASS: Component 3b — Chart is LineChart (0.05 pts)")
                comp3_score += 0.05
            else:
                print(f"FAIL: Component 3b — Expected LineChart, got {type(chart).__name__}")

            # Check chart title contains 'burndown' or 'sprint'
            chart_title_text = None
            try:
                if chart.title:
                    # Title can be a string or Title object
                    if isinstance(chart.title, str):
                        chart_title_text = chart.title
                    else:
                        # Extract text from Title object
                        for p in chart.title.tx.rich.paragraphs:
                            for r in p.r:
                                if r.t:
                                    chart_title_text = (chart_title_text or '') + r.t
            except Exception:
                pass

            if chart_title_text and 'burndown' in chart_title_text.lower():
                print(f"PASS: Component 3c — Chart title: '{chart_title_text}' (0.10 pts)")
                comp3_score += 0.10
            elif chart_title_text and 'sprint' in chart_title_text.lower():
                print(f"PARTIAL: Component 3c — Chart title: '{chart_title_text}' (0.05 pts)")
                comp3_score += 0.05
            else:
                print(f"FAIL: Component 3c — Expected title containing 'burndown', found: {chart_title_text}")

            # Check chart has at least 2 series (ideal line + actual)
            if len(chart.series) >= 2:
                print(f"PASS: Component 3d — Chart has {len(chart.series)} series (0.05 pts)")
                comp3_score += 0.05
            else:
                print(f"FAIL: Component 3d — Expected >=2 series, found {len(chart.series)}")

            total_score += comp3_score
            print(f"  Component 3 total: {comp3_score}/0.35")
        else:
            print(f"FAIL: Component 3 — No charts found in 'Sprint' sheet")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved LibreOffice state
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
