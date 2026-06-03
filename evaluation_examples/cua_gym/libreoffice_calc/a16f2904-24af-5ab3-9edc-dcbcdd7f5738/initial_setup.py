"""
Initial Setup: Build a sprint backlog board view with story point formatting and burndown tracking.
Task ID: calc_gpm_059
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_059'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sprint"

    # --- Title Row: Merge A1:J1 ---
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "Sprint 14 Backlog - Team Phoenix"
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="FF006666", end_color="FF006666", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Row 3: Headers ---
    headers = ['Story ID', 'User Story', 'Points', 'Assignee', 'Status',
               'Priority', 'Day 1', 'Day 5', 'Day 10', 'Completion']
    teal_fill = PatternFill(start_color="FF006666", end_color="FF006666", fill_type="solid")
    white_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="000000")
    all_borders = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = white_font
        cell.fill = teal_fill
        cell.alignment = center_align
        cell.border = all_borders

    # --- Rows 4-15: 12 User Stories ---
    stories = [
        ['SP-101', 'As a user, I want to log in with SSO so I can access the dashboard quickly', 5, 'Liam Torres', 'Done', 'Must Have', 5, 2, 0],
        ['SP-102', 'As a user, I want to reset my password via email link', 3, 'Aisha Patel', 'Done', 'Must Have', 3, 1, 0],
        ['SP-103', 'As an admin, I want to view user activity logs for auditing', 8, 'Marcus Chen', 'In Progress', 'Must Have', 8, 6, 4],
        ['SP-104', 'As a user, I want to filter search results by date range', 5, 'Elena Volkov', 'In Review', 'Should Have', 5, 3, 1],
        ['SP-105', 'As a user, I want to export reports as PDF documents', 3, 'James Okafor', 'In Progress', 'Should Have', 3, 2, 2],
        ['SP-106', 'As an admin, I want to bulk import users from CSV files', 8, 'Priya Sharma', 'To Do', 'Could Have', 8, 8, 8],
        ['SP-107', 'As a user, I want real-time notifications for task assignments', 13, 'David Kim', 'Blocked', 'Should Have', 13, 13, 13],
        ['SP-108', 'As a user, I want to customize my dashboard layout', 5, 'Sofia Reyes', 'Done', 'Could Have', 5, 3, 0],
        ['SP-109', 'As a user, I want to set recurring calendar events', 2, 'Nathan Brooks', 'In Review', 'Should Have', 2, 1, 1],
        ['SP-110', 'As an admin, I want role-based access control for modules', 13, 'Fatima Al-Hassan', 'In Progress', 'Must Have', 13, 10, 7],
        ['SP-111', 'As a user, I want dark mode support across the application', 1, 'Oliver Grant', 'Done', 'Could Have', 1, 0, 0],
        ['SP-112', 'As a user, I want to attach files to task comments', 2, 'Yuki Tanaka', 'To Do', 'Should Have', 2, 2, 2],
    ]

    for r, row_data in enumerate(stories, 4):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = all_borders
            if c in (1, 3, 4, 5, 6):
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Column widths ---
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 14

    # --- Data Validation: Status dropdown E4:E15 ---
    dv_status = DataValidation(
        type="list",
        formula1='"To Do,In Progress,In Review,Done,Blocked"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_status.error = "Invalid status"
    dv_status.errorTitle = "Error"
    dv_status.prompt = "Select status"
    dv_status.promptTitle = "Status"
    dv_status.add("E4:E15")
    ws.add_data_validation(dv_status)

    # --- Data Validation: Priority dropdown F4:F15 ---
    dv_priority = DataValidation(
        type="list",
        formula1='"Must Have,Should Have,Could Have"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_priority.error = "Invalid priority"
    dv_priority.errorTitle = "Error"
    dv_priority.prompt = "Select priority"
    dv_priority.promptTitle = "Priority"
    dv_priority.add("F4:F15")
    ws.add_data_validation(dv_priority)

    # --- Conditional Formatting: Status column E ---
    # Done = green fill
    ws.conditional_formatting.add("E4:E15",
        FormulaRule(formula=['E4="Done"'],
                    fill=PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")))
    # In Progress = blue fill
    ws.conditional_formatting.add("E4:E15",
        FormulaRule(formula=['E4="In Progress"'],
                    fill=PatternFill(start_color="FF5B9BD5", end_color="FF5B9BD5", fill_type="solid")))
    # In Review = purple fill
    ws.conditional_formatting.add("E4:E15",
        FormulaRule(formula=['E4="In Review"'],
                    fill=PatternFill(start_color="FF7030A0", end_color="FF7030A0", fill_type="solid")))
    # To Do = gray
    ws.conditional_formatting.add("E4:E15",
        FormulaRule(formula=['E4="To Do"'],
                    fill=PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")))
    # Blocked = red fill bold
    ws.conditional_formatting.add("E4:E15",
        FormulaRule(formula=['E4="Blocked"'],
                    fill=PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"),
                    font=Font(bold=True)))

    # --- Conditional Formatting: Points column C ---
    # 13 = red (too large)
    ws.conditional_formatting.add("C4:C15",
        CellIsRule(operator="equal", formula=["13"],
                   fill=PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")))
    # 8 = orange
    ws.conditional_formatting.add("C4:C15",
        CellIsRule(operator="equal", formula=["8"],
                   fill=PatternFill(start_color="FFFF8C00", end_color="FFFF8C00", fill_type="solid")))
    # 5 = yellow
    ws.conditional_formatting.add("C4:C15",
        CellIsRule(operator="equal", formula=["5"],
                   fill=PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")))
    # 1-3 = green
    ws.conditional_formatting.add("C4:C15",
        CellIsRule(operator="between", formula=["1", "3"],
                   fill=PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")))

    # J column (Completion) left EMPTY in initial - task requires adding formulas
    # Row 17 left EMPTY - task requires adding totals
    # No chart - task requires adding burndown chart

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
