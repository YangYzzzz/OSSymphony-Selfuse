"""
Initial Setup: Product Revenue spreadsheet with non-contiguous columns
Task ID: calc_fmb_sum_noncontiguous_named_075
Domain: libreoffice_calc

Creates a 'Product Revenue' sheet with monthly data for 4 products in
non-adjacent columns B, D, F, H. Row 12 contains totals. J12 is empty (target).
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_sum_noncontiguous_named_075'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Product Revenue'

    # --- Headers in row 1 ---
    # Column A: Month, B: Product A, C: (blank), D: Product B,
    # E: (blank), F: Product C, G: (blank), H: Product D, I: Label, J: Total
    header_row = {
        'A': 'Month',
        'B': 'Product A',
        'D': 'Product B',
        'F': 'Product C',
        'H': 'Product D',
        'I': '',
        'J': 'Grand Total',
    }
    header_font = Font(name='Calibri', bold=True, size=11)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font_white = Font(name='Calibri', bold=True, size=11, color='FFFFFFFF')

    for col_letter, val in header_row.items():
        cell = ws[f'{col_letter}1']
        cell.value = val
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Monthly data rows 2-11 (10 months) ---
    months = [
        'Jan 2024', 'Feb 2024', 'Mar 2024', 'Apr 2024', 'May 2024',
        'Jun 2024', 'Jul 2024', 'Aug 2024', 'Sep 2024', 'Oct 2024',
    ]

    # Monthly data: [Product A, Product B, Product C, Product D]
    monthly_data = [
        [21500, 28300, 15200, 34100],
        [22800, 29100, 16400, 35700],
        [24100, 30500, 17100, 36200],
        [25300, 31200, 18300, 37800],
        [26700, 32400, 19500, 39100],
        [27900, 33100, 20200, 40300],
        [28400, 34600, 21100, 41500],
        [29600, 35800, 22400, 42700],
        [30200, 36900, 23600, 43900],
        [50900, 50200, 25900, 64500],
    ]

    # Verify column totals match the spec
    # Product A total should be 287400
    # Product B total should be 342100
    # Product C total should be 198700
    # Product D total should be 415800
    expected_totals = [287400, 342100, 198700, 415800]

    # Calculate current totals and adjust last row if needed
    current_totals = [sum(row[i] for row in monthly_data) for i in range(4)]
    adjustments = [expected_totals[i] - current_totals[i] for i in range(4)]
    # Adjust the last row (row index 9) to match expected totals
    for i in range(4):
        monthly_data[9][i] += adjustments[i]

    # Alternate row fill for readability
    light_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    data_cols = ['B', 'D', 'F', 'H']

    for row_idx, (month, values) in enumerate(zip(months, monthly_data), 2):
        use_fill = (row_idx % 2 == 0)
        row_fill = light_fill if use_fill else None

        # Month label in column A
        ws[f'A{row_idx}'] = month
        ws[f'A{row_idx}'].alignment = Alignment(horizontal='left')
        if row_fill:
            ws[f'A{row_idx}'].fill = row_fill

        # Product values in B, D, F, H
        for col_letter, val in zip(data_cols, values):
            cell = ws[f'{col_letter}{row_idx}']
            cell.value = val
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='right')
            if row_fill:
                cell.fill = row_fill

        # Columns C, E, G are intentionally blank (spacers)

    # --- Row 12: Totals ---
    total_font = Font(name='Calibri', bold=True, size=11)
    total_fill = PatternFill(start_color='FFFFC000', end_color='FFFFC000', fill_type='solid')
    thin = Side(style='thin', color='000000')
    total_border = Border(top=thin, bottom=thin)

    ws['A12'] = 'Total'
    ws['A12'].font = total_font
    ws['A12'].fill = total_fill
    ws['A12'].alignment = Alignment(horizontal='left')
    ws['A12'].border = total_border

    # Place exact total values (not formulas) in the total row
    total_values = {'B': 287400, 'D': 342100, 'F': 198700, 'H': 415800}
    for col_letter, val in total_values.items():
        cell = ws[f'{col_letter}12']
        cell.value = val
        cell.font = total_font
        cell.fill = total_fill
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal='right')
        cell.border = total_border

    # I12 = "Grand Total" label
    ws['I12'] = 'Grand Total'
    ws['I12'].font = total_font
    ws['I12'].fill = total_fill
    ws['I12'].alignment = Alignment(horizontal='right')
    ws['I12'].border = total_border

    # J12 is EMPTY (target cell — agent must enter SUM formula here)
    ws['J12'].fill = total_fill
    ws['J12'].border = total_border
    ws['J12'].number_format = '#,##0'
    ws['J12'].alignment = Alignment(horizontal='right')

    # --- Column widths ---
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 4   # spacer
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 4   # spacer
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 4   # spacer
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 14

    # --- Row 1 height ---
    ws.row_dimensions[1].height = 20

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Product Revenue')
    print(f'  Rows: 12 (1 header + 10 data + 1 total)')
    print(f'  B12=287400, D12=342100, F12=198700, H12=415800')
    print(f'  I12="Grand Total", J12=EMPTY (target)')


create_initial()
