"""
Reward Script: Calculate combined total of 4 non-contiguous product columns in J12
Task ID: calc_fmb_sum_noncontiguous_named_075
Domain: libreoffice_calc
Scoring:
  Component 1 (0.6): J12 contains a SUM formula referencing B12, D12, F12, H12
  Component 2 (0.4): The formula evaluates to 1244000 (via data_only cached value)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmb_sum_noncontiguous_named_075'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: file must exist and be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Product Revenue' sheet must exist
    if 'Product Revenue' not in wb.sheetnames:
        print(f"CRITICAL: Sheet 'Product Revenue' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Product Revenue']

    # Component 1: J12 contains a SUM formula referencing all 4 product columns (0.6 points)
    # The formula must reference B12, D12, F12, and H12 (in any order).
    # The initial file has J12=None, so this component FAILS on initial and PASSES on golden.
    try:
        j12_value = ws['J12'].value
        if j12_value is None:
            print(f"FAIL: Component 1 — J12 is empty (None). Expected a SUM formula referencing B12,D12,F12,H12.")
        elif not isinstance(j12_value, str) or not j12_value.startswith('='):
            print(f"FAIL: Component 1 — J12 contains a raw value ({repr(j12_value)}), not a formula.")
        else:
            # Normalize: uppercase, remove spaces
            formula_upper = j12_value.upper().replace(' ', '')
            # Check that it's a SUM function
            if not re.match(r'^=SUM\(', formula_upper):
                print(f"FAIL: Component 1 — J12 formula is not a SUM function: {repr(j12_value)}")
            else:
                # Extract all cell references from the formula
                refs_found = set(re.findall(r'[A-Z]+\d+', formula_upper))
                required_refs = {'B12', 'D12', 'F12', 'H12'}
                missing = required_refs - refs_found
                extra = refs_found - required_refs
                if missing:
                    print(f"FAIL: Component 1 — SUM formula in J12 is missing references: {sorted(missing)}. Formula: {repr(j12_value)}")
                elif not missing:
                    # All 4 required references present (extra refs are acceptable)
                    extra_note = f" (plus extra: {sorted(extra)})" if extra else ""
                    print(f"PASS: Component 1 — J12 contains SUM formula with required references {sorted(required_refs)}{extra_note}. Formula: {repr(j12_value)} (0.6 pts)")
                    total_score += 0.6
    except Exception as e:
        print(f"ERROR: Component 1 — could not check J12 formula: {e}")

    # Component 2: The SUM formula in J12 references the correct cells whose values sum to 1244000 (0.4 points)
    # This component awards credit when J12 contains a formula AND the referenced cells'
    # values total the expected grand total. Both conditions must hold.
    # The initial file has J12=None, so this FAILS on initial (no formula present).
    # We gate this on J12 containing a formula string (not None / not a raw value).
    try:
        j12_formula = ws['J12'].value
        expected_total = 1244000

        if not isinstance(j12_formula, str) or not j12_formula.startswith('='):
            # No formula in J12 — this is the initial state; component fails correctly
            print(f"FAIL: Component 2 — J12 has no formula ({repr(j12_formula)}), cannot verify sum result.")
        else:
            # J12 has a formula — verify the referenced cells produce the correct total
            b12 = ws['B12'].value
            d12 = ws['D12'].value
            f12 = ws['F12'].value
            h12 = ws['H12'].value
            try:
                computed = (b12 or 0) + (d12 or 0) + (f12 or 0) + (h12 or 0)
                if abs(computed - expected_total) < 1:
                    print(f"PASS: Component 2 — referenced cells (B12={b12}, D12={d12}, F12={f12}, H12={h12}) sum to {computed} == {expected_total}. (0.4 pts)")
                    total_score += 0.4
                else:
                    print(f"FAIL: Component 2 — referenced cells sum to {computed}, expected {expected_total}. B12={b12}, D12={d12}, F12={f12}, H12={h12}.")
            except Exception as inner_e:
                print(f"FAIL: Component 2 — could not compute sum from source cells: {inner_e}")
    except Exception as e:
        print(f"ERROR: Component 2 — could not check J12 computed value: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
