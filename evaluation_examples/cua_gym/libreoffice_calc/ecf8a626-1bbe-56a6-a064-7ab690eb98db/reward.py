"""
Reward Script: HR Re-Onboarding Tracker Setup
Task ID: calc_hr_re_onboarding_tracker_075
Domain: libreoffice_calc

Scoring:
  Component 1 (0.25): DATEDIF formulas in E2:E22 (months away calculation)
  Component 2 (0.25): IFS formulas in F2:F22 (re-onboarding duration logic)
  Component 3 (0.20): IF formulas in G2:G22 with date number format (plan end date)
  Component 4 (0.20): Data validation dropdown in H2:H22 (Not Started/In Progress/Completed)
  Component 5 (0.10): Conditional formatting on A2:H22 (green fill when Status=Completed)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_re_onboarding_tracker_075'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the Re-Onboarding sheet exists
    if 'Re-Onboarding' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Re-Onboarding' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Re-Onboarding']

    # Component 1: DATEDIF formulas in E2:E22 (0.25 points)
    # Task requires =DATEDIF(C_row, D_row, "M") for each row 2-22
    # Initial file has all None values in column E - this check fails on initial
    try:
        datedif_count = 0
        datedif_correct = 0
        for row in range(2, 23):  # rows 2 through 22 inclusive
            cell_val = ws.cell(row=row, column=5).value  # column E
            if cell_val is not None:
                datedif_count += 1
                val_str = str(cell_val).upper().replace(' ', '')
                # Check for DATEDIF formula pattern referencing Cx, Dx with "M"
                if ('DATEDIF' in val_str and
                        f'C{row}' in str(cell_val) and
                        f'D{row}' in str(cell_val) and
                        '"M"' in str(cell_val)):
                    datedif_correct += 1

        if datedif_correct == 21:
            print(f"PASS: Component 1 — All 21 DATEDIF formulas present in E2:E22 (0.25 pts)")
            total_score += 0.25
        elif datedif_correct > 0:
            partial = round(0.25 * datedif_correct / 21, 4)
            print(f"PARTIAL: Component 1 — {datedif_correct}/21 DATEDIF formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No DATEDIF formulas found in E2:E22 (found {datedif_count} non-None values)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: IFS formulas in F2:F22 (0.25 points)
    # Task requires =IFS(E_row<3,"1 Week",E_row<6,"2 Weeks",TRUE,"4 Weeks")
    # Initial file has all None values in column F - this check fails on initial
    try:
        ifs_count = 0
        ifs_correct = 0
        for row in range(2, 23):  # rows 2 through 22 inclusive
            cell_val = ws.cell(row=row, column=6).value  # column F
            if cell_val is not None:
                ifs_count += 1
                val_str = str(cell_val).upper().replace(' ', '')
                # Check IFS formula with 3 branches: <3 -> 1 Week, <6 -> 2 Weeks, TRUE -> 4 Weeks
                if ('IFS' in val_str and
                        f'E{row}<3' in val_str.replace(' ', '') and
                        '1WEEK' in val_str.replace('"', '').replace(' ', '') and
                        '2WEEKS' in val_str.replace('"', '').replace(' ', '') and
                        '4WEEKS' in val_str.replace('"', '').replace(' ', '')):
                    ifs_correct += 1

        if ifs_correct == 21:
            print(f"PASS: Component 2 — All 21 IFS formulas present in F2:F22 (0.25 pts)")
            total_score += 0.25
        elif ifs_correct > 0:
            partial = round(0.25 * ifs_correct / 21, 4)
            print(f"PARTIAL: Component 2 — {ifs_correct}/21 IFS formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No IFS formulas found in F2:F22 (found {ifs_count} non-None values)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: IF formulas in G2:G22 with date number format (0.20 points)
    # Task requires =IF(F_row="1 Week", D_row+7, IF(F_row="2 Weeks", D_row+14, D_row+28))
    # and number format should be a date format
    # Initial file has all None values in column G - this check fails on initial
    try:
        if_count = 0
        if_correct = 0
        date_format_ok = False
        for row in range(2, 23):  # rows 2 through 22 inclusive
            cell_val = ws.cell(row=row, column=7).value  # column G
            if cell_val is not None:
                if_count += 1
                val_str = str(cell_val)
                val_upper = val_str.upper().replace(' ', '')
                # Check IF formula referencing F_row and D_row with +7/+14/+28
                if (val_upper.startswith('=IF') and
                        f'F{row}' in val_str and
                        f'D{row}' in val_str and
                        '+7' in val_str and
                        '+14' in val_str and
                        '+28' in val_str):
                    if_correct += 1
                    # Check date format on first passing row
                    if not date_format_ok:
                        fmt = ws.cell(row=row, column=7).number_format
                        if fmt and fmt.lower() not in ['general', '@', '']:
                            # Any non-General format suggests date formatting applied
                            date_format_ok = True

        if if_correct == 21 and date_format_ok:
            print(f"PASS: Component 3 — All 21 IF formulas in G2:G22 with date format (0.20 pts)")
            total_score += 0.20
        elif if_correct == 21:
            print(f"PARTIAL: Component 3 — All 21 IF formulas present but missing date format (0.15 pts)")
            total_score += 0.15
        elif if_correct > 0:
            partial = round(0.20 * if_correct / 21, 4)
            print(f"PARTIAL: Component 3 — {if_correct}/21 IF formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No IF formulas found in G2:G22 (found {if_count} non-None values)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Data validation dropdown in H2:H22 (0.20 points)
    # Task requires dropdown with options: Not Started, In Progress, Completed
    # Initial file has no data validations - this check fails on initial
    try:
        validations = ws.data_validations.dataValidation
        dropdown_found = False
        covers_h_range = False
        correct_options = False

        for dv in validations:
            if dv.type == 'list':
                # Check if the validation covers the H2:H22 range
                sqref_str = str(dv.sqref)
                if 'H2' in sqref_str or 'H' in sqref_str:
                    dropdown_found = True
                    # Check it covers rows 2-22 (at minimum)
                    if 'H2:H22' in sqref_str or ('H2' in sqref_str and 'H22' in sqref_str):
                        covers_h_range = True
                    # Check the dropdown options
                    formula = str(dv.formula1) if dv.formula1 else ''
                    formula_clean = formula.replace('"', '').strip()
                    if ('Not Started' in formula_clean and
                            'In Progress' in formula_clean and
                            'Completed' in formula_clean):
                        correct_options = True

        if dropdown_found and covers_h_range and correct_options:
            print(f"PASS: Component 4 — Data validation dropdown H2:H22 with correct options (0.20 pts)")
            total_score += 0.20
        elif dropdown_found and correct_options:
            print(f"PARTIAL: Component 4 — Dropdown with correct options but wrong range (0.10 pts)")
            total_score += 0.10
        elif dropdown_found:
            print(f"PARTIAL: Component 4 — Dropdown found on H column but options incorrect (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4 — No data validation dropdown found for H column")
            if validations:
                print(f"  Found {len(validations)} validations: " +
                      ", ".join([str(dv.sqref) for dv in validations]))
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Conditional formatting on A2:H22 with green fill when Status=Completed (0.10 points)
    # Task requires formula-based rule: $H2="Completed" -> background #70AD47 (green = FF70AD47)
    # Initial file has no conditional formatting - this check fails on initial
    try:
        cf_rules = ws.conditional_formatting
        cf_found = False
        cf_correct_range = False
        cf_correct_fill = False

        for cf_range, cf_list in cf_rules._cf_rules.items():
            cf_range_str = str(cf_range)
            for rule in cf_list:
                # Look for formula-based rule referencing H column and "Completed"
                if rule.formula:
                    formula_str = str(rule.formula).upper()
                    if 'COMPLETED' in formula_str and 'H' in formula_str:
                        cf_found = True
                        # Check if it covers A2:H22
                        if 'A2' in cf_range_str and 'H22' in cf_range_str:
                            cf_correct_range = True
                        # Check fill color is green #70AD47 (ARGB: FF70AD47)
                        if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                            try:
                                fg_rgb = rule.dxf.fill.fgColor.rgb
                                # Accept FF70AD47 (standard) or 0070AD47 (without alpha)
                                if fg_rgb and '70AD47' in fg_rgb.upper():
                                    cf_correct_fill = True
                            except Exception:
                                pass

        if cf_found and cf_correct_range and cf_correct_fill:
            print(f"PASS: Component 5 — Conditional formatting on A2:H22 with green fill #70AD47 (0.10 pts)")
            total_score += 0.10
        elif cf_found and cf_correct_range:
            print(f"PARTIAL: Component 5 — Conditional formatting on correct range but fill color wrong (0.05 pts)")
            total_score += 0.05
        elif cf_found:
            print(f"PARTIAL: Component 5 — Conditional formatting found but wrong range or missing fill (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — No conditional formatting with 'Completed' formula found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {round(final_score, 4)}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
