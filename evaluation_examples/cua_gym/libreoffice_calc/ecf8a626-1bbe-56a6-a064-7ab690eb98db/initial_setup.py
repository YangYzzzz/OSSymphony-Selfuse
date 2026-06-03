"""
Initial Setup: HR Re-Onboarding Tracker
Task ID: calc_hr_re_onboarding_tracker_075
Domain: libreoffice_calc

Creates a spreadsheet with employee re-onboarding data.
Sheet 'Re-Onboarding' has 21 employees with columns:
  A=Emp ID, B=Name, C=Leave Start, D=Return Date, E=Months Away (empty),
  F=Re-Onboarding Duration (empty), G=Plan End Date (empty), H=Status (empty)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_re_onboarding_tracker_075'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Re-Onboarding'

    # --- Headers (Row 1) ---
    headers = ['Emp ID', 'Name', 'Leave Start', 'Return Date',
               'Months Away', 'Re-Onboarding Duration', 'Plan End Date', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # --- Employee data: realistic names, varying leave lengths ---
    # Dates as date objects for proper Excel date handling
    employees = [
        ('EMP-1021', 'Sarah Chen',        date(2024, 2, 10), date(2025, 3, 3)),   # ~13 months
        ('EMP-1034', 'Marcus Johnson',    date(2024, 9, 1),  date(2025, 2, 24)),  # ~6 months
        ('EMP-1047', 'Priya Nair',        date(2024, 11, 15),date(2025, 2, 17)), # ~3 months
        ('EMP-1058', 'David Okafor',      date(2024, 7, 20), date(2025, 1, 13)), # ~6 months
        ('EMP-1062', 'Aisha Mohammed',    date(2024, 3, 5),  date(2025, 2, 28)), # ~12 months
        ('EMP-1075', 'Thomas Bergmann',   date(2024, 12, 2), date(2025, 2, 10)), # ~2 months
        ('EMP-1083', 'Yuki Tanaka',       date(2024, 8, 12), date(2025, 2, 5)),  # ~6 months
        ('EMP-1091', 'Elena Vasquez',     date(2024, 5, 25), date(2025, 1, 20)), # ~8 months
        ('EMP-1104', 'James O\'Brien',    date(2024, 10, 7), date(2025, 2, 19)), # ~4 months
        ('EMP-1112', 'Fatima Al-Rashid',  date(2024, 1, 14), date(2025, 3, 1)),  # ~14 months
        ('EMP-1125', 'Carlos Mendoza',    date(2024, 6, 3),  date(2025, 1, 28)), # ~8 months
        ('EMP-1138', 'Natalia Ivanova',   date(2024, 11, 28),date(2025, 2, 24)), # ~3 months
        ('EMP-1146', 'Kevin Park',        date(2024, 4, 18), date(2025, 2, 10)), # ~10 months
        ('EMP-1159', 'Diana Fletcher',    date(2024, 9, 22), date(2025, 3, 3)),  # ~5 months
        ('EMP-1167', 'Omar Hassan',       date(2024, 7, 8),  date(2025, 1, 6)),  # ~6 months
        ('EMP-1174', 'Sophie Martin',     date(2024, 12, 10),date(2025, 2, 17)), # ~2 months
        ('EMP-1182', 'Raj Krishnamurthy', date(2024, 3, 20), date(2025, 1, 15)), # ~10 months
        ('EMP-1195', 'Amara Diallo',      date(2024, 10, 25),date(2025, 2, 27)), # ~4 months
        ('EMP-1203', 'Lucas Schneider',   date(2024, 5, 6),  date(2025, 2, 3)),  # ~9 months
        ('EMP-1217', 'Mei-Ling Zhao',     date(2024, 8, 30), date(2025, 2, 12)), # ~5 months
        ('EMP-1224', 'Andre Williams',    date(2024, 2, 22), date(2025, 3, 5)),  # ~12 months
    ]

    for row_idx, (emp_id, name, leave_start, return_date) in enumerate(employees, 2):
        ws.cell(row=row_idx, column=1, value=emp_id)
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=leave_start)
        ws.cell(row=row_idx, column=4, value=return_date)
        # Columns E (5), F (6), G (7), H (8) are intentionally left empty

    # --- Format date columns ---
    for row_idx in range(2, 23):
        ws.cell(row=row_idx, column=3).number_format = 'yyyy-mm-dd'
        ws.cell(row=row_idx, column=4).number_format = 'yyyy-mm-dd'

    # --- Column widths ---
    col_widths = {
        'A': 12,  # Emp ID
        'B': 22,  # Name
        'C': 14,  # Leave Start
        'D': 14,  # Return Date
        'E': 15,  # Months Away
        'F': 24,  # Re-Onboarding Duration
        'G': 14,  # Plan End Date
        'H': 16,  # Status
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # --- Freeze pane at row 2 ---
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Re-Onboarding')
    print(f'  Rows: 21 employee rows (2-22)')
    print(f'  Columns E,F,G,H: empty (no formulas, no validation, no formatting)')


create_initial()
