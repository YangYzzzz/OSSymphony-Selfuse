"""
Initial Setup: Copy Baseline sheet to create Scenario A and Scenario B
Task ID: calc_ps_088
Domain: libreoffice_calc

Creates a workbook with 'Baseline' (model parameters + formulas) and 'Results' sheets.
The agent must copy 'Baseline' twice to create 'Scenario A' and 'Scenario B'.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_088'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Baseline sheet ---
    ws_base = wb.active
    ws_base.title = 'Baseline'

    # Header styling
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    # Title row
    ws_base.merge_cells('A1:D1')
    ws_base['A1'] = 'Revenue Projection Model - Baseline'
    ws_base['A1'].font = Font(name='Calibri', size=14, bold=True, color='2F5496')
    ws_base['A1'].alignment = Alignment(horizontal='center')

    # Section: Input Parameters
    ws_base['A3'] = 'Input Parameters'
    ws_base['A3'].font = Font(name='Calibri', size=12, bold=True, color='2F5496')

    param_headers = ['Parameter', 'Value', 'Unit', 'Notes']
    for col, h in enumerate(param_headers, 1):
        cell = ws_base.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    params = [
        ['Initial Revenue', 1250000, 'USD', 'FY2025 Q1 starting revenue'],
        ['Growth Rate', 0.085, '%/quarter', 'Projected quarterly growth'],
        ['Operating Margin', 0.32, 'ratio', 'Target operating margin'],
        ['Customer Acquisition Cost', 4500, 'USD', 'Average CAC per customer'],
        ['Monthly Churn Rate', 0.025, 'ratio', 'Expected monthly churn'],
        ['Average Contract Value', 18500, 'USD', 'Annual contract value'],
        ['Sales Cycle Length', 45, 'days', 'Average sales cycle'],
        ['Marketing Spend Ratio', 0.15, 'ratio', 'Marketing as % of revenue'],
        ['Support Cost Per User', 120, 'USD/month', 'Per-user support cost'],
        ['Discount Rate', 0.10, 'annual', 'WACC for NPV calculations'],
    ]

    for r, row_data in enumerate(params, 5):
        for c, val in enumerate(row_data, 1):
            cell = ws_base.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 2 and isinstance(val, float) and val < 1:
                cell.number_format = '0.00%'
            elif c == 2 and isinstance(val, (int, float)) and val >= 100:
                cell.number_format = '#,##0'

    # Section: Quarterly Projections
    ws_base['A16'] = 'Quarterly Projections'
    ws_base['A16'].font = Font(name='Calibri', size=12, bold=True, color='2F5496')

    proj_headers = ['Quarter', 'Revenue', 'Operating Profit', 'Cumulative Revenue']
    for col, h in enumerate(proj_headers, 1):
        cell = ws_base.cell(row=17, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    quarters = ['Q1 2025', 'Q2 2025', 'Q3 2025', 'Q4 2025',
                'Q1 2026', 'Q2 2026', 'Q3 2026', 'Q4 2026']

    for i, q in enumerate(quarters):
        row = 18 + i
        ws_base.cell(row=row, column=1, value=q).border = thin_border
        if i == 0:
            ws_base.cell(row=row, column=2, value='=B5').border = thin_border
        else:
            ws_base.cell(row=row, column=2, value=f'=B{row-1}*(1+B6)').border = thin_border
        ws_base.cell(row=row, column=3, value=f'=B{row}*B7').border = thin_border
        if i == 0:
            ws_base.cell(row=row, column=4, value=f'=B{row}').border = thin_border
        else:
            ws_base.cell(row=row, column=4, value=f'=D{row-1}+B{row}').border = thin_border

        for c in range(2, 5):
            ws_base.cell(row=row, column=c).number_format = '$#,##0'

    # Summary formulas
    ws_base['A27'] = 'Total Revenue'
    ws_base['A27'].font = Font(bold=True)
    ws_base['B27'] = '=SUM(B18:B25)'
    ws_base['B27'].number_format = '$#,##0'
    ws_base['B27'].font = Font(bold=True)
    ws_base['B27'].border = Border(top=Side(style='double', color='000000'),
                                    bottom=Side(style='double', color='000000'))

    ws_base['A28'] = 'Total Operating Profit'
    ws_base['A28'].font = Font(bold=True)
    ws_base['B28'] = '=SUM(C18:C25)'
    ws_base['B28'].number_format = '$#,##0'
    ws_base['B28'].font = Font(bold=True)

    # Column widths
    ws_base.column_dimensions['A'].width = 28
    ws_base.column_dimensions['B'].width = 18
    ws_base.column_dimensions['C'].width = 18
    ws_base.column_dimensions['D'].width = 22

    # --- Results sheet ---
    ws_results = wb.create_sheet('Results')

    ws_results.merge_cells('A1:C1')
    ws_results['A1'] = 'Model Results Summary'
    ws_results['A1'].font = Font(name='Calibri', size=14, bold=True, color='2F5496')
    ws_results['A1'].alignment = Alignment(horizontal='center')

    result_headers = ['Metric', 'Value', 'Status']
    for col, h in enumerate(result_headers, 1):
        cell = ws_results.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    results_data = [
        ['Total 8-Quarter Revenue', "=Baseline!B27", 'On Track'],
        ['Total Operating Profit', "=Baseline!B28", 'On Track'],
        ['Average Quarterly Growth', "=Baseline!B6", 'Target Met'],
        ['Operating Margin', "=Baseline!B7", 'Stable'],
        ['Break-even Quarter', 'Q2 2025', 'Achieved'],
        ['Customer LTV', 52000, 'Above Target'],
        ['Payback Period', '8.5 months', 'Within Range'],
        ['Net Promoter Score', 72, 'Excellent'],
        ['Year-over-Year Growth', 0.38, 'Strong'],
        ['Market Share Estimate', 0.12, 'Growing'],
    ]

    for r, row_data in enumerate(results_data, 4):
        for c, val in enumerate(row_data, 1):
            cell = ws_results.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 2 and isinstance(val, (int, float)) and val >= 100:
                cell.number_format = '#,##0'
            elif c == 2 and isinstance(val, float) and val < 1:
                cell.number_format = '0.00%'

    # Status color coding
    status_colors = {
        'On Track': 'FF92D050',
        'Target Met': 'FF92D050',
        'Stable': 'FF00B0F0',
        'Achieved': 'FF92D050',
        'Above Target': 'FF92D050',
        'Within Range': 'FFFFFF00',
        'Excellent': 'FF92D050',
        'Strong': 'FF92D050',
        'Growing': 'FF00B0F0',
    }
    for r in range(4, 14):
        status = ws_results.cell(row=r, column=3).value
        if status in status_colors:
            ws_results.cell(row=r, column=3).fill = PatternFill(
                start_color=status_colors[status],
                end_color=status_colors[status],
                fill_type='solid'
            )

    ws_results.column_dimensions['A'].width = 28
    ws_results.column_dimensions['B'].width = 18
    ws_results.column_dimensions['C'].width = 15

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
