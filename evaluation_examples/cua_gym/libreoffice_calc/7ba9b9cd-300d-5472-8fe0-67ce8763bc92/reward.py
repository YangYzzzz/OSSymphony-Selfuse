"""
Reward Script: Copy Baseline sheet to Scenario A and Scenario B
Task ID: calc_ps_088
Domain: libreoffice_calc
Scoring:
  Component 1 (0.2): 'Scenario A' sheet exists
  Component 2 (0.2): 'Scenario B' sheet exists
  Component 3 (0.2): Sheet order is [Baseline, Scenario A, Scenario B, Results]
  Component 4 (0.2): 'Scenario A' data matches 'Baseline' data
  Component 5 (0.2): 'Scenario B' data matches 'Baseline' data
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_088'


def persist_app_state(domain):
    """Save any unsaved LibreOffice state before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            import time
            time.sleep(0.8)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def sheets_data_match(ws1, ws2):
    """Compare all cell values between two worksheets."""
    if ws1.max_row != ws2.max_row or ws1.max_column != ws2.max_column:
        return False, f"Dimensions differ: {ws1.max_row}x{ws1.max_column} vs {ws2.max_row}x{ws2.max_column}"
    mismatches = []
    for row in range(1, ws1.max_row + 1):
        for col in range(1, ws1.max_column + 1):
            v1 = ws1.cell(row=row, column=col).value
            v2 = ws2.cell(row=row, column=col).value
            if v1 != v2:
                coord = ws1.cell(row=row, column=col).coordinate
                mismatches.append(f"{coord}: '{v1}' vs '{v2}'")
    if mismatches:
        return False, f"{len(mismatches)} mismatches: {mismatches[:5]}"
    return True, "All cells match"


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    sheet_names = wb.sheetnames
    print(f"INFO: Found sheets: {sheet_names}")

    # Component 1: 'Scenario A' sheet exists (0.2 points)
    try:
        if 'Scenario A' in sheet_names:
            print("PASS: Component 1 — 'Scenario A' sheet exists (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 1 — 'Scenario A' sheet not found in {sheet_names}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: 'Scenario B' sheet exists (0.2 points)
    try:
        if 'Scenario B' in sheet_names:
            print("PASS: Component 2 — 'Scenario B' sheet exists (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 2 — 'Scenario B' sheet not found in {sheet_names}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Sheet order is correct (0.2 points)
    # Expected: Baseline at index 0, Scenario A at 1, Scenario B at 2, Results at 3
    try:
        expected_order = ['Baseline', 'Scenario A', 'Scenario B', 'Results']
        if sheet_names == expected_order:
            print(f"PASS: Component 3 — Sheet order is correct: {sheet_names} (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — Expected order {expected_order}, found {sheet_names}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: 'Scenario A' is a copy of 'Baseline' (0.2 points)
    try:
        if 'Scenario A' in sheet_names and 'Baseline' in sheet_names:
            ws_baseline = wb['Baseline']
            ws_scenario_a = wb['Scenario A']
            match, detail = sheets_data_match(ws_baseline, ws_scenario_a)
            if match:
                print(f"PASS: Component 4 — 'Scenario A' data matches 'Baseline' (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 4 — 'Scenario A' does not match 'Baseline': {detail}")
        else:
            print("FAIL: Component 4 — Required sheets not found for comparison")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: 'Scenario B' is a copy of 'Baseline' (0.2 points)
    try:
        if 'Scenario B' in sheet_names and 'Baseline' in sheet_names:
            ws_baseline = wb['Baseline']
            ws_scenario_b = wb['Scenario B']
            match, detail = sheets_data_match(ws_baseline, ws_scenario_b)
            if match:
                print(f"PASS: Component 5 — 'Scenario B' data matches 'Baseline' (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 5 — 'Scenario B' does not match 'Baseline': {detail}")
        else:
            print("FAIL: Component 5 — Required sheets not found for comparison")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
