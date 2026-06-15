"""
Reward Script: What-if pricing analysis with data table, conditional formatting, and surface chart
Task ID: calc_wf_020
Domain: libreoffice_calc
Scoring:
  Component 1 — Data table populated with correct profit values (0.4 pts)
  Component 2 — Conditional formatting applied to data table range (0.3 pts)
  Component 3 — Chart exists (surface-style) visualizing the data (0.3 pts)
"""

import os
import math

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_020'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            import time
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def expected_profit(price, volume, unit_cost=15, fixed_costs=10000):
    """Calculate expected profit: (Price - Cost) * Volume - Fixed Costs"""
    return (price - unit_cost) * volume - fixed_costs


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Pricing' sheet must exist
    if 'Pricing' not in wb.sheetnames:
        print("FAIL: 'Pricing' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Pricing']

    # =========================================================================
    # Component 1: Data table populated with correct profit values (0.4 points)
    #
    # The initial file has prices in A15:A32 (18-35) and volumes in B14:H14
    # (500-2000), but the data cells B15:H32 are ALL empty.
    # The golden file has these cells filled with profit = (price-15)*vol - 10000.
    # We verify that a substantial number of cells contain the correct values.
    # =========================================================================
    try:
        prices = [18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35]
        volumes = [500, 750, 1000, 1250, 1500, 1750, 2000]

        total_cells = len(prices) * len(volumes)  # 126
        correct_cells = 0
        populated_cells = 0

        for r_idx, price in enumerate(prices):
            for c_idx, volume in enumerate(volumes):
                cell_row = 15 + r_idx
                cell_col = 2 + c_idx  # B=2
                cell_val = ws.cell(row=cell_row, column=cell_col).value

                if cell_val is not None:
                    populated_cells += 1
                    exp = expected_profit(price, volume)
                    try:
                        actual = float(cell_val)
                        if abs(actual - exp) < 1.0:  # tolerance for rounding
                            correct_cells += 1
                    except (ValueError, TypeError):
                        pass

        # Score: need at least 80% of cells populated AND correct
        if populated_cells == 0:
            print(f"FAIL: Component 1 — Data table is empty (0/{total_cells} cells populated)")
        else:
            correctness_ratio = correct_cells / total_cells
            populated_ratio = populated_cells / total_cells
            print(f"  Populated: {populated_cells}/{total_cells} ({populated_ratio:.1%})")
            print(f"  Correct: {correct_cells}/{total_cells} ({correctness_ratio:.1%})")

            if correctness_ratio >= 0.8:
                print(f"PASS: Component 1 — Data table correctly populated ({correct_cells}/{total_cells}) (0.4 pts)")
                total_score += 0.4
            elif correctness_ratio >= 0.5:
                partial = 0.2
                print(f"PARTIAL: Component 1 — Data table partially correct ({correct_cells}/{total_cells}) ({partial} pts)")
                total_score += partial
            elif populated_ratio >= 0.5:
                partial = 0.1
                print(f"PARTIAL: Component 1 — Data table populated but values incorrect ({populated_cells}/{total_cells}) ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 1 — Data table mostly empty or incorrect")

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Conditional formatting applied to data table range (0.3 pts)
    #
    # The initial file has 0 conditional formatting rules.
    # The golden file has conditional formatting on B15:H32 with rules for
    # negative (red-ish) and positive (green-ish) profit values.
    # We check that CF rules exist covering the data table range.
    # =========================================================================
    try:
        cf_rules = list(ws.conditional_formatting)
        cf_count = 0
        negative_rule_count = 0
        positive_rule_count = 0

        for cf in cf_rules:
            for rule in cf.rules:
                cf_count += 1
                # Count negative/lessThan rules (red for losses)
                if rule.type == 'cellIs' and rule.operator in ('lessThan', 'lessThanOrEqual'):
                    negative_rule_count += 1
                # Count positive/greaterThan rules (green for profit)
                if rule.type == 'cellIs' and rule.operator in ('greaterThan', 'greaterThanOrEqual'):
                    positive_rule_count += 1

        if cf_count == 0:
            print("FAIL: Component 2 — No conditional formatting rules found")
        elif negative_rule_count > 0 and positive_rule_count > 0:
            print(f"PASS: Component 2 — Conditional formatting with {cf_count} rules, covers negative and positive profit (0.3 pts)")
            total_score += 0.3
        elif negative_rule_count > 0 or positive_rule_count > 0:
            if negative_rule_count > 0:
                print(f"PARTIAL: Component 2 — Conditional formatting has negative rule only (0.15 pts)")
                total_score += 0.15
            elif positive_rule_count > 0:
                print(f"PARTIAL: Component 2 — Conditional formatting has positive rule only (0.15 pts)")
                total_score += 0.15
        elif cf_count > 0:
            print(f"PARTIAL: Component 2 — {cf_count} CF rules found but no recognizable profit/loss highlighting (0.1 pts)")
            total_score += 0.1

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Chart exists (surface-style) visualizing the data (0.3 pts)
    #
    # The initial file has 0 charts.
    # The golden file has a SurfaceChart3D titled "Profit Surface: Price vs Volume"
    # with 7 series (one per volume column).
    # We check for presence of a chart, with bonus for surface type.
    # =========================================================================
    try:
        # Collect all charts from all sheets
        all_charts = []
        for sn in wb.sheetnames:
            all_charts.extend(wb[sn]._charts)

        if len(all_charts) == 0:
            print("FAIL: Component 3 — No charts found in any sheet")
        else:
            chart = all_charts[0]
            chart_class = chart.__class__.__name__
            is_surface = 'Surface' in chart_class
            series_count = len(chart.series)

            if is_surface and series_count >= 3:
                print(f"PASS: Component 3 — Surface chart with {series_count} series ({chart_class}) (0.3 pts)")
                total_score += 0.3
            elif is_surface:
                print(f"PARTIAL: Component 3 — Surface chart but only {series_count} series (0.25 pts)")
                total_score += 0.25
            elif series_count >= 3:
                print(f"PARTIAL: Component 3 — Chart ({chart_class}) with {series_count} series, not surface type (0.2 pts)")
                total_score += 0.2
            elif not is_surface and series_count < 3:
                print(f"PARTIAL: Component 3 — Chart ({chart_class}) found, not surface, few series (0.15 pts)")
                total_score += 0.15

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
