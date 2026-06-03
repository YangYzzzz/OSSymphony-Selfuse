"""
Initial Setup: What-if pricing analysis spreadsheet
Task ID: calc_wf_020
Domain: libreoffice_calc

Creates a Pricing sheet with base scenario parameters, a profit formula,
and a prepared data table area with price points and volume headers.
The agent must: fill in the two-variable data table, add conditional formatting,
and create a surface/heatmap chart.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_020'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pricing"

    # ── Styling helpers ──
    header_font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    label_font = Font(name="Calibri", size=11, bold=True)
    value_font = Font(name="Calibri", size=11)
    currency_fmt = '$#,##0'
    currency_fmt2 = '$#,##0.00'
    number_fmt = '#,##0'
    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    section_fill = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")

    # ── Title ──
    ws.merge_cells("A1:G1")
    ws["A1"] = "What-If Pricing Analysis"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # ── Base Scenario Parameters (rows 3-7) ──
    ws["A3"] = "Base Scenario Parameters"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws["A3"].fill = PatternFill(start_color="FF2E75B6", end_color="FF2E75B6", fill_type="solid")
    ws.merge_cells("A3:C3")
    for col in range(1, 4):
        ws.cell(row=3, column=col).fill = PatternFill(start_color="FF2E75B6", end_color="FF2E75B6", fill_type="solid")
        ws.cell(row=3, column=col).font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")

    params = [
        ("Unit Cost", 15, currency_fmt),
        ("Fixed Costs", 10000, currency_fmt),
        ("Current Price", 25, currency_fmt),
        ("Current Volume", 1000, number_fmt),
    ]
    for i, (label, value, fmt) in enumerate(params):
        row = 4 + i
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=1).fill = section_fill
        ws.cell(row=row, column=1).border = thin_border
        c = ws.cell(row=row, column=2, value=value)
        c.font = value_font
        c.number_format = fmt
        c.border = thin_border
        c.fill = section_fill

    # Named references (using cell addresses for formulas)
    # B4=Unit Cost=15, B5=Fixed Costs=10000, B6=Current Price=25, B7=Current Volume=1000

    # ── Base Profit Calculation (row 9) ──
    ws["A9"] = "Base Profit Calculation"
    ws["A9"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws["A9"].fill = PatternFill(start_color="FF2E75B6", end_color="FF2E75B6", fill_type="solid")
    ws.merge_cells("A9:C9")
    for col in range(1, 4):
        ws.cell(row=9, column=col).fill = PatternFill(start_color="FF2E75B6", end_color="FF2E75B6", fill_type="solid")
        ws.cell(row=9, column=col).font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")

    ws.cell(row=10, column=1, value="Profit Formula").font = label_font
    ws.cell(row=10, column=1).fill = section_fill
    ws.cell(row=10, column=1).border = thin_border
    ws.cell(row=10, column=2, value="(Price - Cost) * Volume - Fixed Costs").font = Font(name="Calibri", size=11, italic=True)
    ws.cell(row=10, column=2).fill = section_fill
    ws.cell(row=10, column=2).border = thin_border

    ws.cell(row=11, column=1, value="Base Profit").font = label_font
    ws.cell(row=11, column=1).fill = section_fill
    ws.cell(row=11, column=1).border = thin_border
    # =(B6-B4)*B7-B5 = (25-15)*1000-10000 = 0
    c = ws.cell(row=11, column=2, value="=(B6-B4)*B7-B5")
    c.font = Font(name="Calibri", size=11, bold=True)
    c.number_format = currency_fmt2
    c.fill = section_fill
    c.border = thin_border

    # ── Data Table Area (rows 13+) ──
    ws["A13"] = "Two-Variable Data Table: Profit by Price & Volume"
    ws["A13"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws["A13"].fill = PatternFill(start_color="FF2E75B6", end_color="FF2E75B6", fill_type="solid")
    ws.merge_cells("A13:H13")
    for col in range(1, 9):
        ws.cell(row=13, column=col).fill = PatternFill(start_color="FF2E75B6", end_color="FF2E75B6", fill_type="solid")
        ws.cell(row=13, column=col).font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")

    # Volume headers across columns B-H (500, 750, 1000, 1250, 1500, 1750, 2000)
    volumes = [500, 750, 1000, 1250, 1500, 1750, 2000]
    ws.cell(row=14, column=1, value="Price \\ Volume").font = Font(name="Calibri", size=10, bold=True, italic=True)
    ws.cell(row=14, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=14, column=1).fill = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")
    ws.cell(row=14, column=1).border = thin_border

    for j, vol in enumerate(volumes):
        c = ws.cell(row=14, column=2 + j, value=vol)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.number_format = number_fmt
        c.alignment = Alignment(horizontal="center")
        c.fill = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")
        c.border = thin_border

    # Price labels down column A (rows 15-32: $18 to $35)
    prices = list(range(18, 36))  # 18 through 35
    for i, price in enumerate(prices):
        c = ws.cell(row=15 + i, column=1, value=price)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.number_format = currency_fmt
        c.alignment = Alignment(horizontal="center")
        c.fill = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")
        c.border = thin_border

    # Data table cells left EMPTY — the agent must fill these with profit formulas
    # Add light borders to the empty data area for visual guidance
    for i in range(len(prices)):
        for j in range(len(volumes)):
            c = ws.cell(row=15 + i, column=2 + j)
            c.border = thin_border
            c.alignment = Alignment(horizontal="center")

    # ── Column widths ──
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
