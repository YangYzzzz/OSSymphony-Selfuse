"""
Initial Setup: Copy formatting from A2:F2 to A10:F10 without changing values
Task ID: calc_cop_paste_special_003
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_cop_paste_special_003'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'StyledReport'

    # --- Define styles ---
    # Header style (row 2): bold, blue background (#003366), white font, bottom border
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF003366', end_color='FF003366', fill_type='solid')
    header_border_bottom = Border(
        bottom=Side(style='thin', color='000000')
    )
    header_alignment = Alignment(horizontal='center', vertical='center')

    # Normal data style
    normal_font = Font(name='Calibri', size=11, bold=False)
    normal_fill = PatternFill(fill_type=None)  # no fill

    # Column headers (Row 1)
    col_headers = ['Region', 'Q1 Sales', 'Q2 Sales', 'Q3 Sales', 'Q4 Sales', 'Annual Total']
    for col, h in enumerate(col_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', size=11, bold=True)

    # Row 2: styled section label row (bold, blue bg #003366, white font, bottom border)
    row2_values = ['EMEA', 28000, 24500, 19200, 13800, 85500]
    for col, val in enumerate(row2_values, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='FF003366', end_color='FF003366', fill_type='solid')
        cell.border = Border(bottom=Side(style='thin', color='000000'))
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Rows 3-9: data rows (no special formatting)
    data_rows = [
        ['Germany',    7200,  6800,  5100,  3600,  22700],
        ['France',     6500,  5900,  4800,  3200,  20400],
        ['UK',         8100,  7200,  5800,  4100,  25200],
        ['Spain',      3200,  2900,  2200,  1500,   9800],
        ['Italy',      2400,  2100,  1700,  1200,   7400],
        ['Netherlands',2800,  2400,  1900,  1400,   8500],
        ['Others',     5800,  5100,  4200,  2800,  17900],
    ]
    for r, row_data in enumerate(data_rows, 3):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name='Calibri', size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Row 10: summary/totals row — NO special formatting applied
    row10_values = ['Total', 142000, 98500, 76200, 54800, 371500]
    for col, val in enumerate(row10_values, 1):
        cell = ws.cell(row=10, column=col, value=val)
        cell.font = Font(name='Calibri', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set column widths for readability
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14

    # Set row heights
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[10].height = 22

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
