"""
Reward Script: Copy formatting from A2:F2 and apply to A10:F10 without changing values
Task ID: calc_cop_paste_special_003
Domain: libreoffice_calc

Task: Copy the formatting (colors, fonts, borders) from A2:F2 and apply it to A10:F10
      without changing any values in row 10.

Expected changes in golden file (Sheet 'StyledReport'):
- Row 10 (A10:F10): bold=True (was False)
- Row 10 (A10:F10): bg color=FF003366 (was 00000000 - no fill)
- Row 10 (A10:F10): font color=00FFFFFF white (was None)
- Row 10 (A10:F10): bottom border=thin (was None)
- Values A10:F10 remain: Total, 142000, 98500, 76200, 54800, 371500

Scoring:
- Component 1: Bold formatting applied to all 6 cells in A10:F10 (0.25 pts)
- Component 2: Blue background color (FF003366) applied to all 6 cells in A10:F10 (0.35 pts)
- Component 3: White font color applied to all 6 cells in A10:F10 (0.20 pts)
- Component 4: Bottom border (thin) applied to all 6 cells in A10:F10 and values preserved (0.20 pts)
Total: 1.0
"""

import os
import openpyxl
from openpyxl.utils import get_column_letter

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_cop_paste_special_003'
SHEET_NAME = 'StyledReport'

# Expected values in row 10 — these should NOT change
EXPECTED_ROW10_VALUES = ['Total', 142000, 98500, 76200, 54800, 371500]

# Expected formatting from row 2 that should be copied to row 10
# Bold is True (verified by checking cell.font.bold directly)
EXPECTED_BG_COLOR = 'FF003366'   # blue background, 8-char ARGB
EXPECTED_FONT_COLOR = '00FFFFFF'  # white font, 8-char ARGB
EXPECTED_BOTTOM_BORDER = 'thin'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook — fail fast if unreadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the target sheet exists — fail fast if missing
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Pre-check: values in row 10 must be preserved
    # (not a scoring component — this is a constraint that must hold)
    row10_values = [ws.cell(row=10, column=c).value for c in range(1, 7)]
    values_preserved = (row10_values == EXPECTED_ROW10_VALUES)
    if not values_preserved:
        print(f"CRITICAL: Values in row 10 have changed!")
        print(f"  Expected: {EXPECTED_ROW10_VALUES}")
        print(f"  Found:    {row10_values}")
        print("REWARD: 0.0")
        return 0.0

    print(f"PASS: Values in A10:F10 preserved correctly: {row10_values}")

    # Component 1: Bold formatting applied to all 6 cells in A10:F10 (0.25 points)
    # This FAILS on initial (bold=False for all) and PASSES on golden (bold=True for all)
    try:
        bold_cells = []
        non_bold_cells = []
        for col in range(1, 7):
            cell = ws.cell(row=10, column=col)
            coord = f"{get_column_letter(col)}10"
            if cell.font.bold:
                bold_cells.append(coord)
            else:
                non_bold_cells.append(coord)

        if len(bold_cells) == 6:
            print(f"PASS: Component 1 — Bold formatting on all 6 cells in row 10 (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — Expected bold on all 6 cells A10:F10.")
            print(f"  Bold cells: {bold_cells}, Non-bold cells: {non_bold_cells}")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not verify bold: {e}")

    # Component 2: Blue background color (FF003366) applied to all 6 cells in A10:F10 (0.35 points)
    # This FAILS on initial (bg=00000000 / no fill) and PASSES on golden (bg=FF003366)
    try:
        correct_bg_cells = []
        wrong_bg_cells = []
        for col in range(1, 7):
            cell = ws.cell(row=10, column=col)
            coord = f"{get_column_letter(col)}10"
            try:
                actual_bg = cell.fill.fgColor.rgb
            except Exception:
                actual_bg = None
            if actual_bg == EXPECTED_BG_COLOR:
                correct_bg_cells.append(coord)
            else:
                wrong_bg_cells.append(f"{coord}(got={actual_bg})")

        if len(correct_bg_cells) == 6:
            print(f"PASS: Component 2 — Blue background (FF003366) on all 6 cells in row 10 (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 2 — Expected background color FF003366 on all 6 cells A10:F10.")
            print(f"  Correct: {correct_bg_cells}, Wrong: {wrong_bg_cells}")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not verify background color: {e}")

    # Component 3: White font color applied to all 6 cells in A10:F10 (0.20 points)
    # This FAILS on initial (font color=None) and PASSES on golden (font color=00FFFFFF)
    try:
        correct_fc_cells = []
        wrong_fc_cells = []
        for col in range(1, 7):
            cell = ws.cell(row=10, column=col)
            coord = f"{get_column_letter(col)}10"
            try:
                actual_fc = cell.font.color.rgb
            except Exception:
                actual_fc = None
            if actual_fc == EXPECTED_FONT_COLOR:
                correct_fc_cells.append(coord)
            else:
                wrong_fc_cells.append(f"{coord}(got={actual_fc})")

        if len(correct_fc_cells) == 6:
            print(f"PASS: Component 3 — White font color (00FFFFFF) on all 6 cells in row 10 (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 3 — Expected white font color 00FFFFFF on all 6 cells A10:F10.")
            print(f"  Correct: {correct_fc_cells}, Wrong: {wrong_fc_cells}")
    except Exception as e:
        print(f"ERROR: Component 3 — Could not verify font color: {e}")

    # Component 4: Bottom border (thin) applied to all 6 cells in A10:F10 (0.20 points)
    # This FAILS on initial (bottom border=None) and PASSES on golden (bottom border=thin)
    try:
        correct_border_cells = []
        wrong_border_cells = []
        for col in range(1, 7):
            cell = ws.cell(row=10, column=col)
            coord = f"{get_column_letter(col)}10"
            b = cell.border
            bottom_style = b.bottom.style if b and b.bottom else None
            if bottom_style == EXPECTED_BOTTOM_BORDER:
                correct_border_cells.append(coord)
            else:
                wrong_border_cells.append(f"{coord}(got={bottom_style})")

        if len(correct_border_cells) == 6:
            print(f"PASS: Component 4 — Bottom border (thin) on all 6 cells in row 10 (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 4 — Expected bottom border 'thin' on all 6 cells A10:F10.")
            print(f"  Correct: {correct_border_cells}, Wrong: {wrong_border_cells}")
    except Exception as e:
        print(f"ERROR: Component 4 — Could not verify bottom border: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
