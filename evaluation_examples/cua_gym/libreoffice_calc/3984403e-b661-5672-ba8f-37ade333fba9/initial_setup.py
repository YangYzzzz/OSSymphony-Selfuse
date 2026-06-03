"""
Initial Setup: Student enrollment data spreadsheet
Task ID: osworld_calc_pivot_multi_styled_015
Domain: libreoffice_calc

Creates a workbook with student enrollment data in Sheet1.
Sheet2 does NOT exist. The agent must create Sheet2 with pivot tables and styled header.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_pivot_multi_styled_015'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Enrollment Data ---
    ws1 = wb.active
    ws1.title = 'Enrollment Data'

    # Headers
    headers = ['Student ID', 'Name', 'Faculty', 'Year Level', 'Enrollment Status', 'GPA']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # Style headers
    header_font = Font(bold=True, name='Calibri', size=11)
    header_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
    for col in range(1, 7):
        cell = ws1.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Realistic student enrollment data (30 rows)
    data = [
        ['S10001', 'Emma Thornton',      'Engineering',        1, 'Active',     3.85],
        ['S10002', 'Liam Nakamura',       'Business',           2, 'Active',     3.42],
        ['S10003', 'Sophia Okonkwo',      'Medicine',           3, 'Active',     3.91],
        ['S10004', 'Noah Patel',          'Arts',               4, 'Active',     3.15],
        ['S10005', 'Isabella Fernandez',  'Engineering',        2, 'Active',     3.67],
        ['S10006', 'James Kowalski',      'Law',                1, 'Active',     3.55],
        ['S10007', 'Mia Andersen',        'Business',           3, 'Inactive',   2.88],
        ['S10008', 'Benjamin Tremblay',   'Medicine',           4, 'Graduated',  3.78],
        ['S10009', 'Charlotte Liu',       'Arts',               1, 'Active',     3.22],
        ['S10010', 'Elijah Osei',         'Engineering',        3, 'Active',     3.49],
        ['S10011', 'Amelia Vasquez',      'Law',                2, 'Graduated',  3.70],
        ['S10012', 'Oliver Bergstrom',    'Business',           4, 'Active',     3.33],
        ['S10013', 'Harper Kim',          'Medicine',           1, 'Active',     3.95],
        ['S10014', 'Lucas Moreau',        'Engineering',        2, 'Inactive',   2.65],
        ['S10015', 'Evelyn Sato',         'Arts',               3, 'Active',     3.10],
        ['S10016', 'Aiden Murphy',        'Law',                4, 'Active',     3.62],
        ['S10017', 'Abigail Hassan',      'Business',           1, 'Active',     3.77],
        ['S10018', 'Jackson Petrov',      'Engineering',        2, 'Graduated',  3.88],
        ['S10019', 'Emily Nguyen',        'Medicine',           3, 'Active',     3.44],
        ['S10020', 'Carter Williams',     'Arts',               4, 'Inactive',   2.71],
        ['S10021', 'Scarlett Thompson',   'Engineering',        1, 'Active',     3.59],
        ['S10022', 'Sebastian Tanaka',    'Law',                2, 'Active',     3.38],
        ['S10023', 'Victoria Okonkwo',    'Business',           3, 'Graduated',  3.82],
        ['S10024', 'Henry Castellano',    'Medicine',           4, 'Active',     3.66],
        ['S10025', 'Grace Lindqvist',     'Arts',               1, 'Active',     3.29],
        ['S10026', 'Michael Chen',        'Engineering',        2, 'Active',     3.74],
        ['S10027', 'Chloe Robinson',      'Law',                3, 'Inactive',   2.93],
        ['S10028', 'Alexander Okafor',    'Business',           4, 'Active',     3.51],
        ['S10029', 'Zoe Martinez',        'Medicine',           1, 'Active',     3.87],
        ['S10030', 'Ryan Sullivan',       'Arts',               2, 'Graduated',  3.40],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Column widths
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 16
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 20
    ws1.column_dimensions['F'].width = 8

    # Freeze header row
    ws1.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open LibreOffice Calc with the initial file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
