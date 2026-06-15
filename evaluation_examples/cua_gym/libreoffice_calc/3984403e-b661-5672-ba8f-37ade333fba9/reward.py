"""
Reward Script: Create pivot tables in Sheet2 with styled merged header
Task ID: osworld_calc_pivot_multi_styled_015
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Sheet2 exists with merged header "Enrollment Summary Report"
                      with blue background and bold white font
  Component 2 (0.25): Pivot table by Faculty has correct counts
  Component 3 (0.25): Pivot table by Year Level has correct counts
  Component 4 (0.20): Pivot table by Enrollment Status has correct counts
  Total: 1.0
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_pivot_multi_styled_015'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: Sheet2 must exist
    if 'Sheet2' not in wb.sheetnames:
        print("FAIL: Sheet2 does not exist in the workbook")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws2 = wb['Sheet2']

    # Component 1: Merged header "Enrollment Summary Report" with blue background
    #              and bold white font (0.30 points)
    # The task requires a merged header at the top of Sheet2 with:
    #   - Text "Enrollment Summary Report"
    #   - Blue background (FF4472C4 or similar blue ARGB)
    #   - Bold white font
    #   - Cell A1 merged across columns
    try:
        header_cell = ws2['A1']
        header_value = header_cell.value

        # Check that A1 contains the header text
        has_header_text = (
            header_value is not None and
            str(header_value).strip() == 'Enrollment Summary Report'
        )

        # Check that the cell is part of a merge range
        has_merge = any(
            str(mr).startswith('A1:') for mr in ws2.merged_cells.ranges
        )

        # Check blue background: fgColor should be a blue (contains "4472C4" or similar blue hex)
        try:
            bg_rgb = header_cell.fill.fgColor.rgb
            # Accept any shade of blue (last 6 chars should indicate blue):
            # Blue means B channel > R and G channels significantly
            # "FF4472C4" is the standard blue used in the golden file
            has_blue_bg = (
                bg_rgb is not None and
                len(bg_rgb) >= 6 and
                header_cell.fill.fill_type == 'solid' and
                (
                    bg_rgb.upper() == 'FF4472C4' or
                    # More flexible: check blue component is dominant
                    (len(bg_rgb) == 8 and int(bg_rgb[6:8], 16) > int(bg_rgb[2:4], 16) and
                     int(bg_rgb[6:8], 16) > int(bg_rgb[4:6], 16))
                )
            )
        except Exception:
            has_blue_bg = False

        # Check bold font
        has_bold = header_cell.font.bold == True

        # Check white font
        try:
            font_rgb = header_cell.font.color.rgb
            has_white_font = (
                font_rgb is not None and
                font_rgb.upper() in ('FFFFFFFF', '00FFFFFF', 'FFFFFF')
            )
        except Exception:
            has_white_font = False

        if has_header_text and has_merge and has_blue_bg and has_bold and has_white_font:
            print(f"PASS: Component 1 — Merged header 'Enrollment Summary Report' with blue bg, bold white font (0.30 pts)")
            total_score += 0.30
        else:
            reasons = []
            if not has_header_text:
                reasons.append(f"header text wrong (found: {repr(header_value)})")
            if not has_merge:
                reasons.append(f"cell A1 not merged (ranges: {ws2.merged_cells.ranges})")
            if not has_blue_bg:
                try:
                    reasons.append(f"background not blue (fgColor.rgb: {header_cell.fill.fgColor.rgb}, fill_type: {header_cell.fill.fill_type})")
                except Exception:
                    reasons.append("background not blue (could not read fill color)")
            if not has_bold:
                reasons.append(f"font not bold (bold: {header_cell.font.bold})")
            if not has_white_font:
                try:
                    reasons.append(f"font not white (color.rgb: {header_cell.font.color.rgb})")
                except Exception:
                    reasons.append("font not white (could not read font color)")
            print(f"FAIL: Component 1 — {'; '.join(reasons)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Pivot table by Faculty has correct counts (0.25 points)
    # Expected: Engineering=7, Business=6, Medicine=6, Arts=6, Law=5
    try:
        expected_faculty = {
            'Engineering': 7,
            'Business': 6,
            'Medicine': 6,
            'Arts': 6,
            'Law': 5,
        }

        # Search Sheet2 for a faculty pivot table section
        # Look for Faculty/Count header and then the data rows
        faculty_found = {}
        faculty_header_row = None

        for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row):
            for cell in row:
                if cell.value == 'Faculty':
                    faculty_header_row = cell.row
                    break
            if faculty_header_row:
                break

        if faculty_header_row:
            # Read rows after the header
            for r in range(faculty_header_row + 1, faculty_header_row + 20):
                if r > ws2.max_row:
                    break
                name_cell = ws2.cell(row=r, column=1)
                count_cell = ws2.cell(row=r, column=2)
                name_val = name_cell.value
                count_val = count_cell.value
                if name_val is None:
                    break
                if str(name_val).strip() in expected_faculty:
                    faculty_found[str(name_val).strip()] = count_val

        # Check if all faculty entries match expected
        if faculty_found and all(
            faculty_found.get(fac) == cnt
            for fac, cnt in expected_faculty.items()
        ) and len(faculty_found) == len(expected_faculty):
            print(f"PASS: Component 2 — Faculty pivot table correct: {faculty_found} (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — Faculty pivot table incorrect. Found: {faculty_found}, Expected: {expected_faculty}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Pivot table by Year Level has correct counts (0.25 points)
    # Expected: Year 1=8, Year 2=8, Year 3=7, Year 4=7
    try:
        expected_year_level = {1: 8, 2: 8, 3: 7, 4: 7}

        year_found = {}
        year_header_row = None

        for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row):
            for cell in row:
                if cell.value == 'Year Level':
                    year_header_row = cell.row
                    break
            if year_header_row:
                break

        if year_header_row:
            for r in range(year_header_row + 1, year_header_row + 20):
                if r > ws2.max_row:
                    break
                level_cell = ws2.cell(row=r, column=1)
                count_cell = ws2.cell(row=r, column=2)
                level_val = level_cell.value
                count_val = count_cell.value
                if level_val is None:
                    break
                try:
                    level_int = int(level_val)
                    if level_int in expected_year_level:
                        year_found[level_int] = count_val
                except (ValueError, TypeError):
                    break

        if year_found and all(
            year_found.get(yr) == cnt
            for yr, cnt in expected_year_level.items()
        ) and len(year_found) == len(expected_year_level):
            print(f"PASS: Component 3 — Year Level pivot table correct: {year_found} (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 3 — Year Level pivot table incorrect. Found: {year_found}, Expected: {expected_year_level}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Pivot table by Enrollment Status has correct counts (0.20 points)
    # Expected: Active=20, Graduated=6, Inactive=4
    try:
        expected_status = {
            'Active': 20,
            'Graduated': 6,
            'Inactive': 4,
        }

        status_found = {}
        status_header_row = None

        for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row):
            for cell in row:
                if cell.value == 'Enrollment Status':
                    status_header_row = cell.row
                    break
            if status_header_row:
                break

        if status_header_row:
            for r in range(status_header_row + 1, status_header_row + 20):
                if r > ws2.max_row:
                    break
                stat_cell = ws2.cell(row=r, column=1)
                count_cell = ws2.cell(row=r, column=2)
                stat_val = stat_cell.value
                count_val = count_cell.value
                if stat_val is None:
                    break
                if str(stat_val).strip() in expected_status:
                    status_found[str(stat_val).strip()] = count_val

        if status_found and all(
            status_found.get(stat) == cnt
            for stat, cnt in expected_status.items()
        ) and len(status_found) == len(expected_status):
            print(f"PASS: Component 4 — Enrollment Status pivot table correct: {status_found} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 4 — Enrollment Status pivot table incorrect. Found: {status_found}, Expected: {expected_status}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the VM
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
