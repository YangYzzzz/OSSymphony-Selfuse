"""
Reward Script: Merge across rows 1 and 2 of columns A through F separately
Task ID: calc_cop_merge_004
Domain: libreoffice_calc

Scoring:
  Component 1: A1:F1 merged as a single row 1 region AND A1 content preserved (0.5 pts)
  Component 2: A2:F2 merged as a single row 2 region AND A2 content preserved (0.4 pts)
  Component 3: Exactly 2 separate merged regions (NOT a combined A1:F2 block) (0.1 pts)
  Total: 1.0
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_cop_merge_004'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check the 'Invoice' sheet exists
    if 'Invoice' not in wb.sheetnames:
        print("FAIL: Sheet 'Invoice' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Invoice']

    # Get all merged cell ranges
    merged_ranges = list(ws.merged_cells.ranges)
    range_strs = [str(r) for r in merged_ranges]

    # Component 1: A1:F1 is a merged region (row 1 merge) AND A1 content preserved (0.5 points)
    # This FAILS on initial (no merged cells exist) and PASSES on golden (A1:F1 merged with content intact)
    try:
        a1f1_merged = 'A1:F1' in range_strs
        if a1f1_merged:
            # Verify B1 through F1 are MergedCell instances (confirming the merge spans to F)
            b1_is_merged = isinstance(ws['B1'], MergedCell)
            f1_is_merged = isinstance(ws['F1'], MergedCell)
            # Verify content in A1 is preserved
            a1_value = ws['A1'].value
            a1_content_ok = a1_value is not None and str(a1_value).strip() == 'ACME Corporation'
            if b1_is_merged and f1_is_merged and a1_content_ok:
                print(f"PASS: Component 1 — A1:F1 merged as single row, A1='{a1_value}' preserved (0.5 pts)")
                total_score += 0.5
            else:
                print(f"FAIL: Component 1 — A1:F1 range present but check failed: "
                      f"B1_merged={b1_is_merged}, F1_merged={f1_is_merged}, A1='{a1_value}'")
        else:
            print(f"FAIL: Component 1 — A1:F1 merged region not found. Ranges present: {range_strs}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: A2:F2 is a merged region (row 2 merge) AND A2 content preserved (0.4 points)
    # This FAILS on initial (no merged cells exist) and PASSES on golden (A2:F2 merged with content intact)
    try:
        a2f2_merged = 'A2:F2' in range_strs
        if a2f2_merged:
            # Verify B2 through F2 are MergedCell instances (confirming the merge spans to F)
            b2_is_merged = isinstance(ws['B2'], MergedCell)
            f2_is_merged = isinstance(ws['F2'], MergedCell)
            # Verify content in A2 is preserved
            a2_value = ws['A2'].value
            a2_content_ok = a2_value is not None and str(a2_value).strip() == 'Invoice #2025-0047'
            if b2_is_merged and f2_is_merged and a2_content_ok:
                print(f"PASS: Component 2 — A2:F2 merged as single row, A2='{a2_value}' preserved (0.4 pts)")
                total_score += 0.4
            else:
                print(f"FAIL: Component 2 — A2:F2 range present but check failed: "
                      f"B2_merged={b2_is_merged}, F2_merged={f2_is_merged}, A2='{a2_value}'")
        else:
            print(f"FAIL: Component 2 — A2:F2 merged region not found. Ranges present: {range_strs}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Two separate merged regions exist (NOT one combined A1:F2 block) (0.1 points)
    # The task requires two INDEPENDENT merged rows, not one big block spanning both rows.
    # This FAILS on initial (no merged cells at all) and PASSES on golden (two separate ranges).
    try:
        a1f2_combined = 'A1:F2' in range_strs
        if a1f2_combined:
            print(f"FAIL: Component 3 — Found forbidden combined A1:F2 block instead of two separate row merges")
        else:
            has_row1 = 'A1:F1' in range_strs
            has_row2 = 'A2:F2' in range_strs
            if has_row1 and has_row2:
                print(f"PASS: Component 3 — Two separate merged regions (A1:F1 and A2:F2), NOT a combined A1:F2 block (0.1 pts)")
                total_score += 0.1
            else:
                print(f"FAIL: Component 3 — Expected both A1:F1 and A2:F2 to be separate. Found: {range_strs}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
