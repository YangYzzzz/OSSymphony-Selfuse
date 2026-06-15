"""
Initial Setup: Merge rows 1 and 2 across columns A-F in Invoice spreadsheet
Task ID: calc_cop_merge_004
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_cop_merge_004'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Invoice ---
    ws = wb.active
    ws.title = 'Invoice'

    # Row 1: Company name in A1, B1:F1 empty (NO merge)
    ws['A1'] = 'ACME Corporation'

    # Row 2: Invoice number in A2, B2:F2 empty (NO merge)
    ws['A2'] = 'Invoice #2025-0047'

    # Row 3: Headers
    headers = ['Item', 'Description', 'Qty', 'Unit Price', 'Discount', 'Total']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)

    # Style headers with bold
    for col in range(1, 7):
        cell = ws.cell(row=3, column=col)
        cell.font = Font(bold=True)

    # Rows 4-15: 12 realistic invoice line items
    line_items = [
        ['001', 'Professional Consulting Services',  10, 450.00, 0.05, 4275.00],
        ['002', 'Software License — Enterprise Tier', 5, 1200.00, 0.10, 5400.00],
        ['003', 'Hardware: Dell XPS Workstation',     3, 2350.00, 0.00, 7050.00],
        ['004', 'Network Switch — 48-Port Managed',   2, 890.00,  0.00, 1780.00],
        ['005', 'Technical Support Contract (Annual)',1, 3600.00, 0.15, 3060.00],
        ['006', 'Training Workshop — 2 Days',         4, 800.00,  0.00, 3200.00],
        ['007', 'Printer Toner Cartridges (Set of 4)',6, 120.00,  0.05, 684.00],
        ['008', 'UPS Battery Backup Unit',            2, 340.00,  0.00, 680.00],
        ['009', 'Cloud Storage Subscription (1 yr)',  3, 550.00,  0.10, 1485.00],
        ['010', 'Ergonomic Office Chair — Premium',   8, 395.00,  0.05, 2998.00],
        ['011', 'Standing Desk Converter',            4, 280.00,  0.00, 1120.00],
        ['012', 'Webcam HD 1080p — Conference Room',  5, 175.00,  0.10, 787.50],
    ]

    for r, row_data in enumerate(line_items, 4):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Column widths for readability
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 6
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12

    # NO merged cells in initial file

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('  - Sheet: Invoice')
    print('  - A1: ACME Corporation (no merge)')
    print('  - A2: Invoice #2025-0047 (no merge)')
    print('  - Row 3: headers (Item, Description, Qty, Unit Price, Discount, Total)')
    print('  - Rows 4-15: 12 realistic line items')
    print('  - No merged cells')


create_initial()
