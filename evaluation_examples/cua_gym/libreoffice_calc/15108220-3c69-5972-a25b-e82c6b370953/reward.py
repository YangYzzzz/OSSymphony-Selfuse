"""
Reward Script: Fix #DIV/0! errors in percentage change formulas
Task ID: calc_tbl_011
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): All D2:D20 formulas use IF(An=0,0,...) pattern
  Component 2 (0.3): Zero-base rows (5,11,17) specifically have IF-guarded formulas
  Component 3 (0.2): All formulas reference correct row numbers
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_011'

# Rows where A column is 0 (the ones that cause #DIV/0!)
ZERO_BASE_ROWS = [5, 11, 17]


def normalize_formula(f):
    """Normalize formula for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ""
    return f.upper().replace(" ", "")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Component 1: All D2:D20 formulas use IF(An=0,0,...) pattern (0.5 points)
    # This checks that ALL 19 formulas have been converted to IF-guarded versions.
    # On initial_env, NONE have IF guards, so this will score 0.
    try:
        if_count = 0
        total_cells = 19  # D2 through D20
        for r in range(2, 21):
            cell_val = ws.cell(row=r, column=4).value
            norm = normalize_formula(cell_val)
            # Check for IF pattern: =IF(An=0,0,(Bn-An)/An)
            # Accept variations like =IF(A5=0,0,(B5-A5)/A5)
            if norm.startswith("=IF("):
                if_count += 1

        if if_count == total_cells:
            print(f"PASS: Component 1 — All {total_cells} formulas use IF guard ({if_count}/{total_cells}) (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — Only {if_count}/{total_cells} formulas use IF guard")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Zero-base rows (5, 11, 17) specifically have IF-guarded formulas (0.3 points)
    # These are the critical rows that produce #DIV/0! in the initial state.
    # On initial_env, these rows have =(Bn-An)/An (no IF), so this will score 0.
    try:
        zero_fixed = 0
        for r in ZERO_BASE_ROWS:
            cell_val = ws.cell(row=r, column=4).value
            norm = normalize_formula(cell_val)
            # Must contain IF and a check for zero: IF(An=0,...)
            expected_pattern = f"=IF(A{r}=0,0,(B{r}-A{r})/A{r})"
            expected_norm = normalize_formula(expected_pattern)
            if norm == expected_norm:
                zero_fixed += 1
                print(f"  Row {r}: PASS — formula is {cell_val}")
            else:
                print(f"  Row {r}: FAIL — expected {expected_pattern}, found {cell_val}")

        if zero_fixed == len(ZERO_BASE_ROWS):
            print(f"PASS: Component 2 — All {len(ZERO_BASE_ROWS)} zero-base rows have correct IF formula (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — Only {zero_fixed}/{len(ZERO_BASE_ROWS)} zero-base rows fixed")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: All formulas reference correct row numbers (0.2 points)
    # Verify each formula in D{r} references A{r} and B{r} (not shifted rows).
    # On initial_env, the formulas do reference correct rows but lack IF, so the
    # specific check here is that the IF-guarded formula references the correct row.
    # This component only awards points if the formula BOTH has IF AND references correct row.
    try:
        correct_refs = 0
        for r in range(2, 21):
            cell_val = ws.cell(row=r, column=4).value
            norm = normalize_formula(cell_val)
            expected = normalize_formula(f"=IF(A{r}=0,0,(B{r}-A{r})/A{r})")
            if norm == expected:
                correct_refs += 1

        if correct_refs == 19:
            print(f"PASS: Component 3 — All 19 formulas reference correct rows with IF guard (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — Only {correct_refs}/19 formulas match exact expected pattern")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved LibreOffice state
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
