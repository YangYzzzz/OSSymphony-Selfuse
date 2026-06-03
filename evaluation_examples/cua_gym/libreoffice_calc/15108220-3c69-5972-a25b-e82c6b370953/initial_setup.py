"""
Initial Setup: Fix #DIV/0! errors in percentage change formulas
Task ID: calc_tbl_011
Domain: libreoffice_calc

Creates a spreadsheet with Base Value (A), New Value (B), Item (C), and
% Change (D) columns.  D2:D20 use =(Bn-An)/An which causes #DIV/0! when
An=0. Rows 5, 11, and 17 have An=0.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_011'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Performance"

    # --- Headers ---
    headers = ["Base Value", "New Value", "Item", "% Change"]
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Data rows (19 rows: 2..20) ---
    # Rows 5, 11, 17 (1-indexed) have base value = 0 to trigger #DIV/0!
    items_data = [
        # (base_value, new_value, item_name)
        (1200,  1350,  "Q1 Widget Sales"),          # row 2
        (8500,  9100,  "Eastern Region Revenue"),    # row 3
        (340,   280,   "Customer Complaints"),        # row 4
        (0,     475,   "New Product Launch"),          # row 5 -> #DIV/0!
        (5600,  5600,  "Warehouse Inventory"),         # row 6
        (2200,  2750,  "Marketing Spend"),             # row 7
        (15000, 13800, "Monthly Active Users"),        # row 8
        (780,   920,   "Support Tickets Resolved"),    # row 9
        (4100,  4500,  "Website Conversions"),         # row 10
        (0,     320,   "Pilot Program Enrollment"),    # row 11 -> #DIV/0!
        (6700,  7100,  "Quarterly Profit"),             # row 12
        (190,   150,   "Equipment Downtime (hrs)"),     # row 13
        (3300,  3800,  "Employee Training Hours"),      # row 14
        (11000, 12500, "Social Media Followers"),       # row 15
        (450,   390,   "Defective Units"),              # row 16
        (0,     150,   "Beta Feature Requests"),        # row 17 -> #DIV/0!
        (8900,  9200,  "Partner Transactions"),         # row 18
        (2100,  2100,  "Office Supply Budget"),         # row 19
        (5400,  6000,  "Client Retention Rate"),        # row 20
    ]

    for i, (base_val, new_val, item) in enumerate(items_data):
        row = i + 2
        ws.cell(row=row, column=1, value=base_val)
        ws.cell(row=row, column=2, value=new_val)
        ws.cell(row=row, column=3, value=item)
        # Formula: =(Bn-An)/An  -- causes #DIV/0! when An=0
        ws.cell(row=row, column=4, value=f'=(B{row}-A{row})/A{row}')

    # --- Formatting ---
    # Number format for column D (percentage)
    for row in range(2, 21):
        ws.cell(row=row, column=4).number_format = '0.00%'
        # Number format for A and B (integer with comma)
        ws.cell(row=row, column=1).number_format = '#,##0'
        ws.cell(row=row, column=2).number_format = '#,##0'

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 14

    # Light alternating row shading
    light_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
    for row in range(2, 21):
        if row % 2 == 0:
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = light_fill

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
