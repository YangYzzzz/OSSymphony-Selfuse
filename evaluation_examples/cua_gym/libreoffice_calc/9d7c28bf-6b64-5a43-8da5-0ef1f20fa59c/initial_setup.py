"""
Initial Setup: Waterfall-style income bridge from gross revenue to net income
Task ID: calc_fin_gross_profit_waterfall_033
Domain: libreoffice_calc

Creates the pre-task state:
- Sheet 'Waterfall' with title in A1, labels in A3:A12, values in B3:B12
- Column C is EMPTY (formulas to be added by agent)
- No conditional formatting, no borders, no bold subtotals, no merged cells, no currency format
"""

import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_gross_profit_waterfall_033'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Waterfall ---
    ws = wb.active
    ws.title = 'Waterfall'

    # A1: Title (not merged, not formatted — agent will merge and format)
    ws['A1'] = 'Income Bridge — FY2024'

    # Column A labels (A3:A12)
    labels = [
        'Gross Revenue',
        'Returns & Allowances',
        'Net Revenue',
        'COGS',
        'Gross Profit',
        'SG&A',
        'R&D',
        'Operating Income',
        'Interest & Other',
        'Net Income',
    ]
    for i, label in enumerate(labels):
        ws.cell(row=3 + i, column=1, value=label)

    # Column B values (B3:B12) — realistic FY2024 financial figures
    # Positive = revenue/income lines, Negative = deductions/expenses
    values = [
        4_820_000,    # Gross Revenue
        -215_000,     # Returns & Allowances (negative)
        4_605_000,    # Net Revenue (= Gross - Returns)
        -2_180_000,   # COGS (negative)
        2_425_000,    # Gross Profit (= Net Revenue - COGS)
        -892_000,     # SG&A (negative)
        -314_000,     # R&D (negative)
        1_219_000,    # Operating Income (= Gross Profit - SG&A - R&D)
        -87_500,      # Interest & Other (negative)
        1_131_500,    # Net Income (= Operating Income + Interest & Other)
    ]
    for i, val in enumerate(values):
        ws.cell(row=3 + i, column=2, value=val)

    # Column C is intentionally left empty — agent will fill with formulas

    # Set reasonable column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: Waterfall')
    print('A1: Title (unmerged, plain)')
    print('A3:A12: Line item labels')
    print('B3:B12: Financial values (raw, no currency format)')
    print('C3:C12: Empty (agent will add formulas)')


create_initial()
