"""
Reward Script: Waterfall-style income bridge setup
Task ID: calc_fin_gross_profit_waterfall_033
Domain: libreoffice_calc
Scoring:
  Component 1: Column C formulas (=B3..=B12) filled in   — 0.25
  Component 2: Bold on key subtotal rows C5, C7, C10, C12 — 0.20
  Component 3: Conditional formatting B3:B12 (green/>0, red/<0) — 0.25
  Component 4: Thick outer border around A3:C12            — 0.15
  Component 5: Currency format B3:B12 + A1:C1 merged + A1 bold size 14 — 0.15
  Total: 1.0
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_gross_profit_waterfall_033'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Waterfall' sheet must exist
    if 'Waterfall' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Waterfall' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Waterfall']

    # Component 1: Column C formulas — C3:C12 each contains =Bx (0.25 points)
    # In the initial file, column C is entirely empty. The task asks to fill C3:C12
    # with references to column B (e.g., C3=B3, C4=B4, ..., C12=B12).
    try:
        formulas_ok = 0
        expected_formulas = {
            3: '=B3', 4: '=B4', 5: '=B5', 6: '=B6', 7: '=B7',
            8: '=B8', 9: '=B9', 10: '=B10', 11: '=B11', 12: '=B12'
        }
        for row, expected in expected_formulas.items():
            val = ws.cell(row=row, column=3).value
            if val is not None and str(val).strip().upper().replace(' ', '') == expected.upper().replace(' ', ''):
                formulas_ok += 1

        if formulas_ok == 10:
            print(f"PASS: Component 1 — All 10 column C formulas present (=B3..=B12) (0.25 pts)")
            total_score += 0.25
        elif formulas_ok >= 5:
            partial = 0.12
            print(f"PARTIAL: Component 1 — {formulas_ok}/10 column C formulas present ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {formulas_ok}/10 column C formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Bold formatting on key subtotal rows in column C
    # C5 (Net Revenue), C7 (Gross Profit), C10 (Operating Income), C12 (Net Income) must be bold.
    # In the initial file, column C is empty/unformatted.
    try:
        bold_rows = [5, 7, 10, 12]
        bold_count = 0
        for row in bold_rows:
            cell = ws.cell(row=row, column=3)
            if cell.font.bold is True:
                bold_count += 1

        if bold_count == 4:
            print(f"PASS: Component 2 — All 4 subtotal rows bold in column C (C5, C7, C10, C12) (0.20 pts)")
            total_score += 0.20
        elif bold_count >= 2:
            partial = 0.10
            print(f"PARTIAL: Component 2 — {bold_count}/4 subtotal rows bold in column C ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {bold_count}/4 subtotal rows bold in column C")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Conditional formatting on B3:B12
    # >0 → green font (FF00B050), <0 → red font (FFFF0000)
    # Initial file has NO conditional formatting.
    try:
        cf_rules_all = ws.conditional_formatting._cf_rules
        green_rule_count = 0
        red_rule_count = 0

        for cf_range, rules in cf_rules_all.items():
            for rule in rules:
                rule_type = rule.type
                rule_operator = getattr(rule, 'operator', None)
                rule_formula = getattr(rule, 'formula', None)

                if rule_type == 'cellIs' and rule_formula and '0' in str(rule_formula):
                    if hasattr(rule, 'dxf') and rule.dxf is not None:
                        dxf = rule.dxf
                        if dxf.font is not None:
                            try:
                                c = dxf.font.color
                                c_type = c.type
                                if c_type == 'rgb':
                                    c_rgb = c.rgb
                                    # Count green rule (FF00B050 or alternate greens) for >0
                                    if rule_operator == 'greaterThan' and c_rgb in ('FF00B050', 'FF00FF00', '0000FF00'):
                                        green_rule_count += 1
                                    # Count red rule (FFFF0000 or alternate reds) for <0
                                    elif rule_operator == 'lessThan' and (c_rgb == 'FFFF0000' or c_rgb.upper().endswith('FF0000')):
                                        red_rule_count += 1
                            except Exception:
                                pass

        if green_rule_count >= 1 and red_rule_count >= 1:
            print(f"PASS: Component 3 — Conditional formatting: green for >0, red for <0 on B3:B12 (0.25 pts)")
            total_score += 0.25
        elif green_rule_count >= 1 or red_rule_count >= 1:
            partial = 0.12
            which = 'green' if green_rule_count >= 1 else 'red'
            print(f"PARTIAL: Component 3 — Only {which} CF rule found ({partial} pts)")
            total_score += partial
        else:
            # Check if ANY conditional formatting rules exist (partial credit)
            cf_count = sum(len(r) for r in cf_rules_all.values())
            if cf_count > 0:
                print(f"PARTIAL: Component 3 — CF rules exist ({cf_count}) but colors don't match expected (0.08 pts)")
                total_score += 0.08
            else:
                print(f"FAIL: Component 3 — No conditional formatting found on B3:B12")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Thick outer border around A3:C12 (0.15 points)
    # Initial file has no borders. The task requires a thick border around the whole range.
    # Check the 4 outer edges: left of col A (rows 3-12), right of col C (rows 3-12),
    # top of row 3 (cols A-C), bottom of row 12 (cols A-C).
    try:
        border_checks = {
            'left_col_A': [ws.cell(row=r, column=1).border.left.style for r in range(3, 13)
                           if ws.cell(row=r, column=1).border.left],
            'right_col_C': [ws.cell(row=r, column=3).border.right.style for r in range(3, 13)
                            if ws.cell(row=r, column=3).border.right and
                            not isinstance(ws.cell(row=r, column=3), MergedCell)],
            'top_row3': [ws.cell(row=3, column=c).border.top.style for c in range(1, 4)
                         if ws.cell(row=3, column=c).border.top],
            'bottom_row12': [ws.cell(row=12, column=c).border.bottom.style for c in range(1, 4)
                             if ws.cell(row=12, column=c).border.bottom],
        }

        thick_edges = 0
        for edge_name, styles in border_checks.items():
            if any(s in ('thick', 'medium', 'thin') for s in styles if s):
                thick_edges += 1

        # Check specifically for thick style on corner cells
        corner_thick = 0
        # Top-left corner: A3 should have thick left and thick top
        a3 = ws['A3']
        if a3.border.left and a3.border.left.style == 'thick':
            corner_thick += 1
        if a3.border.top and a3.border.top.style == 'thick':
            corner_thick += 1
        # Bottom-right corner: C12 should have thick right and thick bottom
        c12 = ws['C12']
        if c12.border.right and c12.border.right.style == 'thick':
            corner_thick += 1
        if c12.border.bottom and c12.border.bottom.style == 'thick':
            corner_thick += 1

        if corner_thick >= 4:
            print(f"PASS: Component 4 — Thick outer border around A3:C12 (0.15 pts)")
            total_score += 0.15
        elif corner_thick >= 2 or thick_edges >= 2:
            partial = 0.07
            print(f"PARTIAL: Component 4 — Partial border detected (corner_thick={corner_thick}, thick_edges={thick_edges}) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — No thick border detected around A3:C12 (corner_thick={corner_thick})")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Currency format on B3:B12 + A1:C1 merged + A1 bold size 14 (0.15 points)
    # Initial file: B3:B12 has 'General' format, A1 is not merged and not bold size 14.
    try:
        sub_score = 0.0

        # Check currency format on B3:B12 (0.05 pts)
        currency_ok = 0
        for row in range(3, 13):
            cell = ws.cell(row=row, column=2)
            nf = cell.number_format
            if nf and ('$' in nf or '#,##0' in nf):
                currency_ok += 1

        if currency_ok == 10:
            sub_score += 0.05
            print(f"  PASS: C5a — Currency format on B3:B12")
        elif currency_ok >= 5:
            sub_score += 0.02
            print(f"  PARTIAL: C5a — {currency_ok}/10 cells have currency format")
        else:
            print(f"  FAIL: C5a — Only {currency_ok}/10 cells have currency format")

        # Check A1:C1 merged (0.05 pts)
        merged_ranges = [str(mc) for mc in ws.merged_cells.ranges]
        a1_merged = any('A1:C1' in mr or mr == 'A1:C1' for mr in merged_ranges)
        if a1_merged:
            sub_score += 0.05
            print(f"  PASS: C5b — A1:C1 is merged")
        else:
            print(f"  FAIL: C5b — A1:C1 not merged (found: {merged_ranges})")

        # Check A1 bold and size 14 (0.05 pts)
        a1 = ws['A1']
        a1_bold = a1.font.bold is True
        a1_size = a1.font.size
        if a1_bold and a1_size is not None and float(a1_size) == 14.0:
            sub_score += 0.05
            print(f"  PASS: C5c — A1 is bold and size 14")
        elif a1_bold:
            sub_score += 0.02
            print(f"  PARTIAL: C5c — A1 is bold but size={a1_size} (expected 14)")
        else:
            print(f"  FAIL: C5c — A1 bold={a1_bold}, size={a1_size}")

        if sub_score > 0:
            print(f"PASS: Component 5 — Sub-score {sub_score}/0.15 pts")
        total_score += sub_score

    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
