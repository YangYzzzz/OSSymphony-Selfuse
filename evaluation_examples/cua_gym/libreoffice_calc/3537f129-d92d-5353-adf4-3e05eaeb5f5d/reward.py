"""
Reward Script: Create a pivot table counting employees per department
Task ID: calc_pivot_004
Domain: libreoffice_calc
Scoring:
  Component 1 (0.2): Pivot/summary sheet exists (not in initial)
  Component 2 (0.2): All 5 department labels present
  Component 3 (0.4): Correct employee counts per department
  Component 4 (0.2): Grand total row equals 85
"""

import os
from collections import Counter

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_004'

# Expected department counts from task context
EXPECTED_COUNTS = {
    'HR': 12,
    'Engineering': 28,
    'Marketing': 15,
    'Sales': 18,
    'Finance': 12,
}
EXPECTED_TOTAL = 85


def get_count_value(ws_pivot, row, col, ws_employees=None):
    """
    Get the count value from a pivot cell. Handles both:
    - Direct numeric values
    - COUNTIF formulas (compute the actual count from Employees sheet)
    """
    cell = ws_pivot.cell(row=row, column=col)
    val = cell.value

    # If it's a number, return it directly
    if isinstance(val, (int, float)):
        return int(val)

    # If it's a COUNTIF formula, compute the count manually from Employees data
    if isinstance(val, str) and 'COUNTIF' in val.upper():
        # Get the department label from column A of the same row
        dept_label = ws_pivot.cell(row=row, column=col - 1).value
        if dept_label and ws_employees:
            count = 0
            for r in range(2, ws_employees.max_row + 1):
                if ws_employees.cell(row=r, column=3).value == dept_label:
                    count += 1
            return count

    # If it's a SUM formula for grand total, compute from Employees
    if isinstance(val, str) and 'SUM' in val.upper():
        if ws_employees:
            count = 0
            for r in range(2, ws_employees.max_row + 1):
                if ws_employees.cell(row=r, column=3).value:
                    count += 1
            return count

    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: A pivot/summary sheet exists beyond 'Employees' (0.2 points)
    # The initial file has only 'Employees'. A new sheet must be added for the pivot.
    try:
        non_employee_sheets = [s for s in wb.sheetnames if s != 'Employees']
        if len(non_employee_sheets) >= 1:
            pivot_sheet_name = non_employee_sheets[0]
            print(f"PASS: Component 1 — Found pivot sheet '{pivot_sheet_name}' (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 1 — No additional sheet found beyond 'Employees'. Sheets: {wb.sheetnames}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    if total_score < 0.1:
        # No pivot sheet found — cannot check further components
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Get references to pivot and employees sheets
    pivot_sheet_name = [s for s in wb.sheetnames if s != 'Employees'][0]
    ws_pivot = wb[pivot_sheet_name]
    ws_employees = wb['Employees'] if 'Employees' in wb.sheetnames else None

    # Build a map of department label -> (row, count_value) from the pivot sheet
    # Scan all rows to find department labels and their corresponding counts
    dept_data = {}
    grand_total_val = None
    count_col = None

    # Find the column layout: look for department names and count values
    for row in range(1, ws_pivot.max_row + 1):
        for col in range(1, ws_pivot.max_column + 1):
            cell_val = ws_pivot.cell(row=row, column=col).value
            if isinstance(cell_val, str) and cell_val in EXPECTED_COUNTS:
                # Found a department label; count should be in the next column
                count_col_candidate = col + 1
                count_val = get_count_value(ws_pivot, row, count_col_candidate, ws_employees)
                dept_data[cell_val] = count_val
                count_col = count_col_candidate
            if isinstance(cell_val, str) and cell_val.lower().replace(' ', '') in ('grandtotal', 'total'):
                # Found grand total label
                gt_col = col + 1
                grand_total_val = get_count_value(ws_pivot, row, gt_col, ws_employees)

    # Component 2: All 5 department labels present (0.2 points)
    try:
        found_depts = set(dept_data.keys())
        expected_depts = set(EXPECTED_COUNTS.keys())
        if found_depts >= expected_depts:
            print(f"PASS: Component 2 — All 5 departments found: {sorted(found_depts)} (0.2 pts)")
            total_score += 0.2
        else:
            missing = expected_depts - found_depts
            print(f"FAIL: Component 2 — Missing departments: {missing}. Found: {sorted(found_depts)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Correct employee counts per department (0.4 points)
    # Each correct count earns 0.08 points (5 departments * 0.08 = 0.4)
    try:
        correct_count = 0
        for dept, expected_val in EXPECTED_COUNTS.items():
            actual_val = dept_data.get(dept)
            if actual_val is not None and actual_val == expected_val:
                print(f"  PASS: {dept} count = {actual_val} (expected {expected_val})")
                correct_count += 1
            else:
                print(f"  FAIL: {dept} count = {actual_val} (expected {expected_val})")

        if correct_count == 5:
            dept_score = 0.4
            print(f"PASS: Component 3 — All 5 department counts correct (0.4 pts)")
            total_score += dept_score
        elif correct_count > 0:
            dept_score = correct_count * 0.08
            print(f"PARTIAL: Component 3 — {correct_count}/5 department counts correct ({dept_score:.2f} pts)")
            total_score += dept_score
        else:
            print(f"FAIL: Component 3 — No department counts correct")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Grand total row equals 85 (0.2 points)
    try:
        if grand_total_val is not None and grand_total_val == EXPECTED_TOTAL:
            print(f"PASS: Component 4 — Grand total = {grand_total_val} (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 4 — Grand total = {grand_total_val} (expected {EXPECTED_TOTAL})")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved LibreOffice state before verification
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
persist_app_state()

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
