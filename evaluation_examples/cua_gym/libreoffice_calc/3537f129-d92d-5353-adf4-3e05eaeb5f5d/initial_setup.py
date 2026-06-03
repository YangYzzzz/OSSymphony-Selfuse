"""
Initial Setup: Create employee database spreadsheet for pivot table task
Task ID: calc_pivot_004
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_004'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Employees'

    # --- Headers ---
    headers = ['EmpID', 'Name', 'Department', 'Title', 'Salary', 'HireDate']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # --- Employee Data ---
    # Department distribution: HR=12, Engineering=28, Marketing=15, Sales=18, Finance=12
    dept_config = [
        ('HR', 12),
        ('Engineering', 28),
        ('Marketing', 15),
        ('Sales', 18),
        ('Finance', 12),
    ]

    first_names = [
        'Sarah', 'Marcus', 'Priya', 'James', 'Wei', 'Aisha', 'Carlos',
        'Emily', 'Dmitri', 'Yuki', 'Michael', 'Fatima', 'Liam', 'Zara',
        'Raj', 'Olivia', 'Hassan', 'Mia', 'Daniel', 'Sofia', 'Kenji',
        'Isabella', 'Andre', 'Nadia', 'Thomas', 'Amara', 'Lucas', 'Elena',
        'Oscar', 'Leila', 'Ryan', 'Aaliya', 'Noah', 'Chloe', 'Victor',
        'Maya', 'Simon', 'Lucia', 'Ethan', 'Nina', 'Gabriel', 'Rose',
        'Adrian', 'Jasmine', 'Felix', 'Hannah', 'Marco', 'Aria', 'David',
        'Samantha', 'Julian', 'Tara', 'Patrick', 'Ines', 'Kevin', 'Diana',
        'Alex', 'Natalia', 'Brian', 'Eva', 'Peter', 'Lena', 'George',
        'Marta', 'Ivan', 'Sasha', 'Henry', 'Julia', 'Frank', 'Vera',
        'Leo', 'Clara', 'Sam', 'Iris', 'Tom', 'Lily', 'Max', 'Grace',
        'Ben', 'Ella', 'Jack', 'Zoe', 'Paul', 'Anna', 'Owen',
    ]

    last_names = [
        'Chen', 'Johnson', 'Patel', 'Williams', 'Zhang', 'Ibrahim', 'Garcia',
        'Davis', 'Petrov', 'Tanaka', 'Brown', 'Al-Rashid', 'Murphy', 'Khan',
        'Sharma', 'Wilson', 'Ahmed', 'Anderson', 'Kim', 'Martinez', 'Sato',
        'Rivera', 'Dubois', 'Okafor', 'Mueller', 'Diallo', 'Johansson', 'Popov',
        'Fernandez', 'Singh', 'O\'Brien', 'Hassan', 'Lee', 'Laurent', 'Nguyen',
        'Costa', 'Berg', 'Romano', 'Campbell', 'Volkov', 'Herrera', 'Fischer',
        'Moreau', 'Park', 'Santos', 'Lindberg', 'Rossi', 'Ali', 'Thompson',
        'Wang', 'Torres', 'Nakamura', 'Scott', 'Gonzalez', 'Hansen', 'Reed',
        'Evans', 'Kowalski', 'Hughes', 'Meyer', 'Fox', 'Bauer', 'Cooper',
        'Cruz', 'Ivanov', 'Novak', 'Grant', 'Larsen', 'Wells', 'Kato',
        'Stone', 'Hart', 'Bell', 'Ross', 'Price', 'Ward', 'Cole', 'Lane',
        'Blake', 'Reid', 'Hunt', 'King', 'Ford', 'Boyd', 'Nash',
    ]

    titles_by_dept = {
        'HR': ['HR Manager', 'Recruiter', 'HR Specialist', 'Benefits Coordinator',
               'HR Analyst', 'Talent Acquisition Lead', 'HR Business Partner',
               'Compensation Analyst', 'Training Coordinator', 'HR Director',
               'Employee Relations Specialist', 'HR Coordinator'],
        'Engineering': ['Software Engineer', 'Senior Software Engineer', 'Staff Engineer',
                        'Engineering Manager', 'DevOps Engineer', 'QA Engineer',
                        'Frontend Developer', 'Backend Developer', 'Data Engineer',
                        'ML Engineer', 'Principal Engineer', 'Tech Lead',
                        'Site Reliability Engineer', 'Security Engineer',
                        'Platform Engineer', 'Mobile Developer', 'Solutions Architect',
                        'Engineering Director', 'Cloud Engineer', 'Systems Engineer',
                        'Full Stack Developer', 'Infrastructure Engineer',
                        'Release Engineer', 'Build Engineer', 'Test Automation Engineer',
                        'Performance Engineer', 'Database Engineer', 'Embedded Engineer'],
        'Marketing': ['Marketing Manager', 'Content Strategist', 'SEO Specialist',
                      'Digital Marketing Analyst', 'Brand Manager', 'Social Media Manager',
                      'Email Marketing Specialist', 'Growth Hacker', 'Marketing Coordinator',
                      'Product Marketing Manager', 'Campaign Manager',
                      'Marketing Director', 'Creative Director', 'Copywriter',
                      'Marketing Analyst'],
        'Sales': ['Account Executive', 'Sales Manager', 'Business Development Rep',
                  'Sales Director', 'Enterprise Account Manager', 'Inside Sales Rep',
                  'Solutions Consultant', 'Sales Operations Analyst', 'Channel Manager',
                  'Regional Sales Manager', 'Sales Engineer', 'Customer Success Manager',
                  'Key Account Manager', 'Sales Coordinator', 'Territory Manager',
                  'Partnership Manager', 'Revenue Analyst', 'Sales Trainer'],
        'Finance': ['Financial Analyst', 'Senior Accountant', 'Finance Manager',
                    'Controller', 'Treasury Analyst', 'Tax Specialist',
                    'Budget Analyst', 'Accounts Payable Specialist', 'CFO',
                    'Internal Auditor', 'Financial Planner', 'Payroll Specialist'],
    }

    salary_ranges = {
        'HR': (55000, 105000),
        'Engineering': (75000, 180000),
        'Marketing': (52000, 120000),
        'Sales': (50000, 145000),
        'Finance': (60000, 130000),
    }

    # Build employee records in department order, then shuffle
    employees = []
    emp_idx = 0
    for dept, count in dept_config:
        for i in range(count):
            emp_id = f'E{emp_idx + 1:03d}'
            name = f'{first_names[emp_idx]} {last_names[emp_idx]}'
            title = titles_by_dept[dept][i % len(titles_by_dept[dept])]
            sal_low, sal_high = salary_ranges[dept]
            salary = round(random.randint(sal_low, sal_high) / 500) * 500
            hire_start = date(2018, 1, 1)
            hire_end = date(2025, 9, 30)
            delta_days = (hire_end - hire_start).days
            hire_date = hire_start + timedelta(days=random.randint(0, delta_days))
            employees.append((emp_id, name, dept, title, salary, hire_date))
            emp_idx += 1

    # Shuffle to mix departments (realistic)
    random.shuffle(employees)

    # Write data rows
    for r, (emp_id, name, dept, title, salary, hire_date) in enumerate(employees, 2):
        ws.cell(row=r, column=1, value=emp_id)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=dept)
        ws.cell(row=r, column=4, value=title)
        cell_sal = ws.cell(row=r, column=5, value=salary)
        cell_sal.number_format = '$#,##0'
        cell_hd = ws.cell(row=r, column=6, value=hire_date)
        cell_hd.number_format = 'yyyy-mm-dd'

    # Column widths for readability
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
