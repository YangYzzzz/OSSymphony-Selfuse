"""
Reward Script: Calculate reorder point for each SKU using ROP = (Avg Daily Demand * Lead Time) + Safety Stock
Task ID: calc_ops_051
Domain: libreoffice_calc
Scoring: 5 formula checks in E2:E6 (0.2 each), verifying both formula structure and computed values
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_051'

# Ground truth: expected formulas and computed values
EXPECTED = {
    'E2': {'formula': '=B2*C2+D2', 'value': 450},
    'E3': {'formula': '=B3*C3+D3', 'value': 570},
    'E4': {'formula': '=B4*C4+D4', 'value': 475},
    'E5': {'formula': '=B5*C5+D5', 'value': 620},
    'E6': {'formula': '=B6*C6+D6', 'value': 350},
}

WEIGHT_PER_CELL = 0.2  # 5 cells * 0.2 = 1.0


def persist_app_state():
    """Save any unsaved LibreOffice changes via Ctrl+S."""
    try:
        import time
        os.environ["DISPLAY"] = ":0"
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(1.0)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def check_formula(val, expected_formula):
    """Check if cell value matches expected formula (case-insensitive, space-stripped)."""
    if not isinstance(val, str):
        return False
    return val.upper().replace(" ", "") == expected_formula.upper().replace(" ", "")


def check_computed_value(val, expected_num, tolerance=0.01):
    """Check if a numeric value matches expected (for data_only mode)."""
    if val is None:
        return False
    try:
        return abs(float(val) - expected_num) <= tolerance
    except (ValueError, TypeError):
        return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook for formula checking
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['ROP'] if 'ROP' in wb.sheetnames else None
    if ws is None:
        print("CRITICAL: Sheet 'ROP' not found")
        print("REWARD: 0.0")
        return 0.0

    # Also load with data_only to get cached computed values
    try:
        wb_data = openpyxl.load_workbook(file_path, data_only=True)
        ws_data = wb_data['ROP']
    except Exception:
        ws_data = None

    for cell_ref, expected in EXPECTED.items():
        # Component: Check formula or computed value in cell (0.2 points each)
        try:
            formula_val = ws[cell_ref].value
            has_formula = check_formula(formula_val, expected['formula'])

            # Also accept if the computed value is correct (agent might use
            # a different but equivalent formula, or a hardcoded correct value)
            data_val = ws_data[cell_ref].value if ws_data else None
            has_correct_value = check_computed_value(data_val, expected['value'])

            # Also check if the cell itself has the correct numeric value
            # (in case the file was saved with computed values)
            has_numeric = check_computed_value(formula_val, expected['value'])

            if has_formula:
                print(f"PASS: {cell_ref} — correct formula '{formula_val}' ({WEIGHT_PER_CELL} pts)")
                total_score += WEIGHT_PER_CELL
            elif has_correct_value:
                print(f"PASS: {cell_ref} — correct computed value {data_val} (expected {expected['value']}) ({WEIGHT_PER_CELL} pts)")
                total_score += WEIGHT_PER_CELL
            elif has_numeric:
                print(f"PASS: {cell_ref} — correct numeric value {formula_val} (expected {expected['value']}) ({WEIGHT_PER_CELL} pts)")
                total_score += WEIGHT_PER_CELL
            else:
                print(f"FAIL: {cell_ref} — expected formula '{expected['formula']}' or value {expected['value']}, found formula='{formula_val}', cached={data_val}")
        except Exception as e:
            print(f"ERROR: {cell_ref} — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state()

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
