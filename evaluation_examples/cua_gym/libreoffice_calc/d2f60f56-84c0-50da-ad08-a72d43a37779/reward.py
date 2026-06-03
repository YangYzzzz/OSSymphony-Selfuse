"""
Reward Script: Create 'to_read' sheet with unread NYT non-fiction books
Task ID: osworld_multi_apps_misc_015
Domain: libreoffice_calc
Scoring:
  Component 1: 'to_read' sheet exists in workbook (0.3 pts)
  Component 2: Headers match expected columns (0.2 pts)
  Component 3: Books are not already in user's read list and have valid ranks/year (0.3 pts)
  Component 4: Rows are sorted by Rank ascending (0.2 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_misc_015'

# Expected headers (same as main sheet)
EXPECTED_HEADERS = ['Rank', 'Title', 'Author', 'Year Published']

def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Load the user's read list from main sheet as precondition
    read_titles = set()
    try:
        if 'my_books' in wb.sheetnames:
            ws_main = wb['my_books']
            for row in ws_main.iter_rows(min_row=2, values_only=True):
                if row[1]:
                    read_titles.add(str(row[1]).strip().lower())
        print(f"INFO: Found {len(read_titles)} books in user's read list")
    except Exception as e:
        print(f"WARN: Could not read main sheet: {e}")

    # Component 1: 'to_read' sheet exists (0.3 points)
    try:
        if 'to_read' in wb.sheetnames:
            print("PASS: Component 1 — 'to_read' sheet exists in workbook (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — 'to_read' sheet not found. Sheets: {wb.sheetnames}")
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws_read = wb['to_read']

    # Component 2: Headers are correct (0.2 points)
    try:
        header_row = [ws_read.cell(row=1, column=c).value for c in range(1, 5)]
        # Normalize headers for comparison
        normalized_headers = [str(h).strip() if h is not None else '' for h in header_row]
        if normalized_headers == EXPECTED_HEADERS:
            print(f"PASS: Component 2 — Headers correct: {header_row} (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 2 — Expected headers {EXPECTED_HEADERS}, found {header_row}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Books are not in user's read list and published before 2024 (0.3 points)
    # Check that at least some books qualify as valid unread books from NYT list
    try:
        to_read_data = []
        for row in ws_read.iter_rows(min_row=2, values_only=True):
            if row[0] is not None or row[1] is not None:
                to_read_data.append(row)

        if len(to_read_data) == 0:
            print("FAIL: Component 3 — 'to_read' sheet has no data rows")
        else:
            # Check each book is not in the read list and year < 2024
            valid_books = 0
            invalid_books = []
            for row in to_read_data:
                rank, title, author, year = row[0], row[1], row[2], row[3]
                title_str = str(title).strip() if title else ''
                title_lower = title_str.lower()

                # Check if already read
                if title_lower in read_titles:
                    invalid_books.append(f"  '{title_str}' is already in user's read list")
                    continue

                # Check year published < 2024
                try:
                    year_int = int(year) if year is not None else 9999
                    if year_int >= 2024:
                        invalid_books.append(f"  '{title_str}' year {year_int} is not before 2024")
                        continue
                except (ValueError, TypeError):
                    invalid_books.append(f"  '{title_str}' has invalid year: {year}")
                    continue

                # Check rank is an integer in reasonable range (1-100)
                try:
                    rank_int = int(rank) if rank is not None else -1
                    if rank_int < 1 or rank_int > 100:
                        invalid_books.append(f"  '{title_str}' has unreasonable rank: {rank_int}")
                        continue
                except (ValueError, TypeError):
                    invalid_books.append(f"  '{title_str}' has invalid rank: {rank}")
                    continue

                valid_books += 1

            total_books = len(to_read_data)
            if invalid_books:
                print(f"WARN: Component 3 — Some books have issues:")
                for issue in invalid_books:
                    print(issue)

            # Award points if most books are valid (at least 75% valid)
            if total_books > 0 and valid_books / total_books >= 0.75:
                print(f"PASS: Component 3 — {valid_books}/{total_books} books valid (unread, pre-2024) (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 3 — Only {valid_books}/{total_books} books are valid. Need >= 75%")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Rows are sorted by Rank ascending (0.2 points)
    try:
        rank_values = []
        for row in ws_read.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                try:
                    rank_values.append(int(row[0]))
                except (ValueError, TypeError):
                    rank_values.append(None)

        # Filter out None values for comparison
        valid_ranks = [r for r in rank_values if r is not None]

        if len(valid_ranks) == 0:
            print("FAIL: Component 4 — No rank values found to check sorting")
        elif valid_ranks == sorted(valid_ranks):
            print(f"PASS: Component 4 — Ranks sorted ascending: {valid_ranks} (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 4 — Ranks NOT sorted ascending. Found: {valid_ranks}, expected: {sorted(valid_ranks)}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in a given env
file_path = f'{WORKDIR}/my_books.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
