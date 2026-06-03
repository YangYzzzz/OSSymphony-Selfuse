"""
Initial Setup: Books read spreadsheet + Chrome with NYT Best Sellers
Task ID: osworld_multi_apps_misc_015
Domain: libreoffice_calc (multi-app: also Chrome)

Creates:
  - /home/user/my_books.xlsx with a 'my_books' sheet listing books the user has read
  - Opens LibreOffice Calc with my_books.xlsx
  - Opens Chrome with nytimes.com/books/best-sellers

The 'my_books' sheet uses columns: Rank, Title, Author, Year Published
The user's read list includes some NYT non-fiction bestsellers (pre-2024) and
various other books, so the agent must identify which Top 20 non-fiction
pre-2024 NYT books are missing and create a 'to_read' sheet.
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_misc_015'
OUTPUT = f'{WORKDIR}/my_books.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Main sheet: my_books ---
    ws = wb.active
    ws.title = 'my_books'

    # Headers matching task spec: Rank, Title, Author, Year Published
    headers = ['Rank', 'Title', 'Author', 'Year Published']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Books the user has already read.
    # This list includes:
    #   - Some NYT non-fiction bestsellers published before 2024 (these should NOT appear in to_read)
    #   - Various fiction and non-fiction books to make it realistic
    #
    # NYT non-fiction pre-2024 books ALREADY READ (so they won't be in to_read):
    #   - Outlive: The Science and Art of Longevity (Peter Attia, 2023)
    #   - The Wager (David Grann, 2023)
    #   - Spare (Prince Harry, 2023)
    #   - Greenlights (Matthew McConaughey, 2020)
    #   - The Body Keeps the Score (Bessel van der Kolk, 2014)
    #
    # The user's list also contains fiction and older non-fiction to round it out.
    # Note: In my_books, Rank represents the user's personal ranking (1=favorite).
    books_read = [
        # Rank, Title, Author, Year Published
        (1,  'Outlive: The Science and Art of Longevity', 'Peter Attia',           2023),
        (2,  'The Wager: A Tale of Shipwreck, Mutiny and Murder', 'David Grann',   2023),
        (3,  'Spare',                                        'Prince Harry',        2023),
        (4,  'Greenlights',                                  'Matthew McConaughey', 2020),
        (5,  'The Body Keeps the Score',                     'Bessel van der Kolk', 2014),
        (6,  'Tomorrow, and Tomorrow, and Tomorrow',         'Gabrielle Zevin',     2022),
        (7,  'Fourth Wing',                                  'Rebecca Yarros',      2023),
        (8,  'Happy Place',                                  'Emily Henry',         2023),
        (9,  'Lessons in Chemistry',                         'Bonnie Garmus',       2022),
        (10, 'The House in the Cerulean Sea',               'TJ Klune',            2020),
        (11, 'Atomic Habits',                                'James Clear',         2018),
        (12, 'Educated',                                     'Tara Westover',       2018),
        (13, 'Braiding Sweetgrass',                          'Robin Wall Kimmerer', 2013),
        (14, 'Hidden Pictures',                              'Jason Rekulak',       2022),
        (15, 'The Light We Carry',                           'Michelle Obama',      2022),
        (16, 'Sea of Tranquility',                           'Emily St. John Mandel', 2022),
        (17, 'Demon Copperhead',                             'Barbara Kingsolver',  2022),
        (18, 'Crying in H Mart',                             'Michelle Zauner',     2021),
        (19, 'All the Light We Cannot See',                  'Anthony Doerr',       2014),
        (20, 'The Midnight Library',                         'Matt Haig',           2020),
        (21, 'Where the Crawdads Sing',                      'Delia Owens',         2018),
        (22, 'Think Again',                                  'Adam Grant',          2021),
        (23, 'The Atlas Six',                                'Olivie Blake',        2022),
        (24, 'Anxious People',                               'Fredrik Backman',     2020),
        (25, 'Project Hail Mary',                            'Andy Weir',           2021),
    ]

    for r, row_data in enumerate(books_read, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open LibreOffice Calc with the spreadsheet
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)

    # Open Chrome with the NYT Best Sellers page
    launch_gui('google-chrome --new-window "https://www.nytimes.com/books/best-sellers"', delay_sec=2.0)

    print('GUI_READY: launched LibreOffice Calc and Chrome with DISPLAY=:0')


create_initial()
