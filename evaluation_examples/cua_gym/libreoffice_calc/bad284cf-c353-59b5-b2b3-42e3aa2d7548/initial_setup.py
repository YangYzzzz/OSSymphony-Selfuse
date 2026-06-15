"""
Initial Setup: Create Data_Freshness spreadsheet with record data and dates
Task ID: calc_gcv_045
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_045'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data_Freshness"

    # --- Headers ---
    headers = ["Record ID", "Source System", "Last Updated", "Value"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Data generation ---
    source_systems = [
        "SAP ERP", "Salesforce CRM", "Oracle DB", "AWS S3 Bucket",
        "Snowflake DW", "HubSpot", "Jira Cloud", "ServiceNow",
        "Azure SQL", "Google BigQuery", "MongoDB Atlas", "PostgreSQL",
        "Tableau Server", "Power BI Service", "Kafka Stream",
    ]

    # Generate 49 rows of data (rows 2-50)
    # Dates spread from 2025-01-01 to 2026-03-28
    start_date = date(2025, 1, 1)
    end_date = date(2026, 3, 28)
    date_range_days = (end_date - start_date).days

    random.seed(42)  # reproducible

    data_rows = []
    for i in range(49):
        record_id = f"REC-{1001 + i}"
        source = source_systems[i % len(source_systems)]
        # Spread dates across the full range with some clustering
        day_offset = random.randint(0, date_range_days)
        last_updated = start_date + timedelta(days=day_offset)
        value = round(random.uniform(10.0, 9999.99), 2)
        data_rows.append((record_id, source, last_updated, value))

    date_format = 'yyyy-mm-dd'
    currency_format = '#,##0.00'

    for r, (rec_id, source, dt, val) in enumerate(data_rows, 2):
        ws.cell(row=r, column=1, value=rec_id).border = thin_border
        ws.cell(row=r, column=2, value=source).border = thin_border
        date_cell = ws.cell(row=r, column=3, value=dt)
        date_cell.number_format = date_format
        date_cell.border = thin_border
        val_cell = ws.cell(row=r, column=4, value=val)
        val_cell.number_format = currency_format
        val_cell.border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14

    # Freeze header row
    ws.freeze_panes = "A2"

    # NO conditional formatting in initial state
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
