"""
Reward Script: Apply alternating tab colors to sheets
Task ID: calc_gsi_060
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): Odd-numbered sheets (1,3,5,7,9) have blue tab color
  Component 2 (0.5): Even-numbered sheets (2,4,6,8,10) have orange tab color
  Each correct sheet contributes 0.1 points.
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_060'

# Acceptable blue values — the golden uses 000000FF but an agent might use
# various blue shades.  We accept common "pure blue" ARGB representations.
BLUE_ACCEPTABLE = {'000000FF', 'FF0000FF', '0000FF'}
# Acceptable orange values — golden uses 00FF8000.
ORANGE_ACCEPTABLE = {'00FF8000', 'FFFF8000', 'FF8000', '00FFA500', 'FFFFA500', 'FFA500'}


def _normalize_rgb(rgb_val):
    """Return the raw rgb string (uppercase, stripped) or None."""
    if rgb_val is None:
        return None
    s = str(rgb_val).strip().upper()
    # openpyxl sometimes returns '00000000' for the default theme placeholder
    if s in ('00000000', ''):
        return None
    return s


def _is_blue(rgb_str):
    """Check if the rgb string is a recognizable blue."""
    if rgb_str is None:
        return False
    # Direct match
    if rgb_str in BLUE_ACCEPTABLE:
        return True
    # Flexible: last 6 chars represent RGB, check blue channel dominant
    hex6 = rgb_str[-6:] if len(rgb_str) >= 6 else rgb_str
    try:
        r = int(hex6[0:2], 16)
        g = int(hex6[2:4], 16)
        b = int(hex6[4:6], 16)
        # Blue dominant: b > 180 and b > r and b > g
        if b >= 180 and b > r and b > g:
            return True
    except (ValueError, IndexError):
        pass
    return False


def _is_orange(rgb_str):
    """Check if the rgb string is a recognizable orange."""
    if rgb_str is None:
        return False
    if rgb_str in ORANGE_ACCEPTABLE:
        return True
    hex6 = rgb_str[-6:] if len(rgb_str) >= 6 else rgb_str
    try:
        r = int(hex6[0:2], 16)
        g = int(hex6[2:4], 16)
        b = int(hex6[4:6], 16)
        # Orange: high red, moderate green, low blue
        if r >= 200 and 80 <= g <= 200 and b <= 80:
            return True
    except (ValueError, IndexError):
        pass
    return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: must have 10 sheets
    if len(wb.sheetnames) < 10:
        print(f"FAIL: Precondition — expected at least 10 sheets, found {len(wb.sheetnames)}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Odd-numbered sheets have blue tabs (0.5 points, 0.1 each)
    odd_sheets = ['Sheet1', 'Sheet3', 'Sheet5', 'Sheet7', 'Sheet9']
    odd_pass = 0
    for sheet_name in odd_sheets:
        try:
            if sheet_name not in wb.sheetnames:
                print(f"FAIL: Component 1 — sheet '{sheet_name}' not found")
                continue
            ws = wb[sheet_name]
            tab_color = ws.sheet_properties.tabColor
            rgb = _normalize_rgb(tab_color.rgb if tab_color else None)
            if _is_blue(rgb):
                print(f"PASS: Component 1 — '{sheet_name}' tab is blue (rgb={rgb})")
                odd_pass += 1
                total_score += 0.1
            else:
                print(f"FAIL: Component 1 — '{sheet_name}' expected blue tab, found rgb={rgb}")
        except Exception as e:
            print(f"ERROR: Component 1 — '{sheet_name}': {e}")

    print(f"Component 1 subtotal: {odd_pass}/5 odd sheets blue ({odd_pass * 0.1:.1f} pts)")

    # Component 2: Even-numbered sheets have orange tabs (0.5 points, 0.1 each)
    even_sheets = ['Sheet2', 'Sheet4', 'Sheet6', 'Sheet8', 'Sheet10']
    even_pass = 0
    for sheet_name in even_sheets:
        try:
            if sheet_name not in wb.sheetnames:
                print(f"FAIL: Component 2 — sheet '{sheet_name}' not found")
                continue
            ws = wb[sheet_name]
            tab_color = ws.sheet_properties.tabColor
            rgb = _normalize_rgb(tab_color.rgb if tab_color else None)
            if _is_orange(rgb):
                print(f"PASS: Component 2 — '{sheet_name}' tab is orange (rgb={rgb})")
                even_pass += 1
                total_score += 0.1
            else:
                print(f"FAIL: Component 2 — '{sheet_name}' expected orange tab, found rgb={rgb}")
        except Exception as e:
            print(f"ERROR: Component 2 — '{sheet_name}': {e}")

    print(f"Component 2 subtotal: {even_pass}/5 even sheets orange ({even_pass * 0.1:.1f} pts)")

    final_score = round(min(total_score, 1.0), 1)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
