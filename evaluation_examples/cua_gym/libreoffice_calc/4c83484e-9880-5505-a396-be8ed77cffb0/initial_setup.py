"""
Initial Setup: Apply alternating tab colors to sheets
Task ID: calc_gsi_060
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_060'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: Sales Overview ---
    ws1 = wb.active
    ws1.title = 'Sheet1'
    headers = ['Product', 'Region', 'Q1 Sales', 'Q2 Sales', 'Q3 Sales', 'Q4 Sales']
    for c, h in enumerate(headers, 1):
        ws1.cell(row=1, column=c, value=h)
        ws1.cell(row=1, column=c).font = Font(bold=True)
    data = [
        ['Laptop Pro 15', 'North America', 245000, 312000, 287000, 356000],
        ['Laptop Pro 15', 'Europe', 198000, 221000, 215000, 267000],
        ['Tablet X10', 'North America', 156000, 178000, 192000, 213000],
        ['Tablet X10', 'Asia Pacific', 203000, 245000, 267000, 298000],
        ['SmartWatch V3', 'North America', 89000, 112000, 134000, 156000],
        ['SmartWatch V3', 'Europe', 67000, 89000, 98000, 121000],
        ['Headphones Elite', 'North America', 45000, 56000, 63000, 78000],
        ['Headphones Elite', 'Asia Pacific', 34000, 42000, 51000, 67000],
        ['Keyboard Mech', 'Europe', 23000, 31000, 28000, 35000],
        ['Keyboard Mech', 'North America', 31000, 38000, 42000, 49000],
        ['Mouse Pro', 'Asia Pacific', 18000, 22000, 26000, 31000],
        ['Mouse Pro', 'Europe', 15000, 19000, 21000, 27000],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # --- Sheet2: Employee Directory ---
    ws2 = wb.create_sheet('Sheet2')
    headers2 = ['Employee ID', 'Name', 'Department', 'Title', 'Hire Date', 'Salary']
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=c, value=h)
        ws2.cell(row=1, column=c).font = Font(bold=True)
    emp_data = [
        ['EMP-001', 'Sarah Chen', 'Engineering', 'Senior Developer', '2021-03-15', 125000],
        ['EMP-002', 'Marcus Johnson', 'Marketing', 'Campaign Manager', '2020-07-01', 92000],
        ['EMP-003', 'Priya Sharma', 'Engineering', 'Tech Lead', '2019-11-20', 145000],
        ['EMP-004', 'James Wilson', 'Finance', 'Financial Analyst', '2022-01-10', 85000],
        ['EMP-005', 'Aisha Okafor', 'HR', 'Recruitment Lead', '2021-06-15', 88000],
        ['EMP-006', 'Carlos Rivera', 'Engineering', 'DevOps Engineer', '2020-09-01', 118000],
        ['EMP-007', 'Emma Thompson', 'Sales', 'Account Executive', '2023-02-14', 78000],
        ['EMP-008', 'Wei Zhang', 'Engineering', 'ML Engineer', '2022-04-01', 135000],
        ['EMP-009', 'Fatima Al-Hassan', 'Product', 'Product Manager', '2021-08-22', 115000],
        ['EMP-010', 'David Kim', 'Finance', 'Controller', '2018-12-01', 155000],
    ]
    for r, row_data in enumerate(emp_data, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    # --- Sheet3: Inventory Tracking ---
    ws3 = wb.create_sheet('Sheet3')
    headers3 = ['SKU', 'Item Name', 'Category', 'In Stock', 'Reorder Point', 'Unit Price']
    for c, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=c, value=h)
        ws3.cell(row=1, column=c).font = Font(bold=True)
    inv_data = [
        ['SKU-1001', 'Wireless Mouse', 'Peripherals', 342, 100, 29.99],
        ['SKU-1002', 'USB-C Hub', 'Accessories', 156, 50, 49.99],
        ['SKU-1003', '27" Monitor', 'Displays', 78, 25, 349.99],
        ['SKU-1004', 'Mechanical Keyboard', 'Peripherals', 215, 75, 89.99],
        ['SKU-1005', 'Webcam HD', 'Peripherals', 412, 150, 59.99],
        ['SKU-1006', 'Desk Lamp LED', 'Office', 189, 60, 34.99],
        ['SKU-1007', 'Standing Desk Mat', 'Office', 95, 30, 44.99],
        ['SKU-1008', 'Cable Management Kit', 'Accessories', 267, 80, 19.99],
        ['SKU-1009', 'Noise-Cancel Headset', 'Audio', 134, 40, 179.99],
        ['SKU-1010', 'Laptop Stand', 'Accessories', 198, 60, 39.99],
    ]
    for r, row_data in enumerate(inv_data, 2):
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r, column=c, value=val)

    # --- Sheet4: Project Timeline ---
    ws4 = wb.create_sheet('Sheet4')
    headers4 = ['Project', 'Phase', 'Start Date', 'End Date', 'Status', 'Budget']
    for c, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=c, value=h)
        ws4.cell(row=1, column=c).font = Font(bold=True)
    proj_data = [
        ['Alpha Launch', 'Planning', '2025-01-10', '2025-02-15', 'Complete', 50000],
        ['Alpha Launch', 'Development', '2025-02-16', '2025-05-30', 'In Progress', 200000],
        ['Alpha Launch', 'Testing', '2025-06-01', '2025-07-15', 'Pending', 75000],
        ['Beta Expansion', 'Research', '2025-03-01', '2025-04-15', 'Complete', 30000],
        ['Beta Expansion', 'Design', '2025-04-16', '2025-06-30', 'In Progress', 85000],
        ['Beta Expansion', 'Implementation', '2025-07-01', '2025-10-31', 'Pending', 320000],
        ['Gamma Migration', 'Assessment', '2025-02-01', '2025-03-15', 'Complete', 25000],
        ['Gamma Migration', 'Execution', '2025-03-16', '2025-08-31', 'In Progress', 150000],
        ['Gamma Migration', 'Validation', '2025-09-01', '2025-10-15', 'Pending', 45000],
        ['Delta Optimization', 'Analysis', '2025-04-01', '2025-05-15', 'In Progress', 35000],
    ]
    for r, row_data in enumerate(proj_data, 2):
        for c, val in enumerate(row_data, 1):
            ws4.cell(row=r, column=c, value=val)

    # --- Sheet5: Customer Feedback ---
    ws5 = wb.create_sheet('Sheet5')
    headers5 = ['Ticket ID', 'Customer', 'Product', 'Rating', 'Category', 'Date']
    for c, h in enumerate(headers5, 1):
        ws5.cell(row=1, column=c, value=h)
        ws5.cell(row=1, column=c).font = Font(bold=True)
    fb_data = [
        ['TKT-2001', 'Global Tech Inc', 'Laptop Pro 15', 4.5, 'Performance', '2025-03-01'],
        ['TKT-2002', 'StartupCo', 'Tablet X10', 3.8, 'Battery Life', '2025-03-05'],
        ['TKT-2003', 'MegaCorp LLC', 'SmartWatch V3', 4.2, 'Design', '2025-03-08'],
        ['TKT-2004', 'Fresh Foods Ltd', 'Headphones Elite', 4.7, 'Sound Quality', '2025-03-10'],
        ['TKT-2005', 'DataStream AG', 'Keyboard Mech', 3.5, 'Durability', '2025-03-12'],
        ['TKT-2006', 'BioHealth Corp', 'Mouse Pro', 4.0, 'Ergonomics', '2025-03-15'],
        ['TKT-2007', 'EduLearn Inc', 'Laptop Pro 15', 4.8, 'Display', '2025-03-18'],
        ['TKT-2008', 'RetailMax', 'Webcam HD', 3.9, 'Image Quality', '2025-03-20'],
        ['TKT-2009', 'TravelStar', 'Noise-Cancel Headset', 4.6, 'Comfort', '2025-03-22'],
        ['TKT-2010', 'FinanceFirst', 'USB-C Hub', 4.1, 'Connectivity', '2025-03-25'],
    ]
    for r, row_data in enumerate(fb_data, 2):
        for c, val in enumerate(row_data, 1):
            ws5.cell(row=r, column=c, value=val)

    # --- Sheet6: Budget Summary ---
    ws6 = wb.create_sheet('Sheet6')
    headers6 = ['Department', 'Q1 Budget', 'Q1 Actual', 'Q2 Budget', 'Q2 Actual', 'Variance']
    for c, h in enumerate(headers6, 1):
        ws6.cell(row=1, column=c, value=h)
        ws6.cell(row=1, column=c).font = Font(bold=True)
    budget_data = [
        ['Engineering', 500000, 487000, 520000, 534000, -14000],
        ['Marketing', 200000, 215000, 220000, 198000, 7000],
        ['Sales', 150000, 142000, 160000, 167000, 1000],
        ['HR', 80000, 78000, 85000, 82000, 5000],
        ['Finance', 100000, 95000, 105000, 101000, 9000],
        ['Product', 180000, 192000, 190000, 185000, -7000],
        ['Operations', 250000, 238000, 260000, 271000, 1000],
        ['Legal', 120000, 115000, 125000, 118000, 12000],
        ['Customer Support', 90000, 94000, 95000, 99000, -8000],
        ['R&D', 350000, 342000, 370000, 381000, -3000],
    ]
    for r, row_data in enumerate(budget_data, 2):
        for c, val in enumerate(row_data, 1):
            ws6.cell(row=r, column=c, value=val)

    # --- Sheet7: Vendor Contacts ---
    ws7 = wb.create_sheet('Sheet7')
    headers7 = ['Vendor ID', 'Company', 'Contact Person', 'Email', 'Phone', 'Contract Expires']
    for c, h in enumerate(headers7, 1):
        ws7.cell(row=1, column=c, value=h)
        ws7.cell(row=1, column=c).font = Font(bold=True)
    vendor_data = [
        ['V-101', 'TechSupply Co', 'Robert Hayes', 'r.hayes@techsupply.com', '555-0101', '2025-12-31'],
        ['V-102', 'CloudNet Services', 'Linda Park', 'l.park@cloudnet.io', '555-0102', '2026-06-30'],
        ['V-103', 'SecurIT Solutions', 'Ahmed Nasser', 'a.nasser@securit.com', '555-0103', '2025-09-15'],
        ['V-104', 'GreenOffice Inc', 'Maria Santos', 'm.santos@greenoffice.com', '555-0104', '2026-03-01'],
        ['V-105', 'DataVault Ltd', 'Chen Wei', 'c.wei@datavault.co', '555-0105', '2025-11-30'],
        ['V-106', 'NetWorks Pro', 'Olga Petrov', 'o.petrov@networks.pro', '555-0106', '2026-01-15'],
        ['V-107', 'PixelPrint Media', 'Jake Morrison', 'j.morrison@pixelprint.com', '555-0107', '2025-08-31'],
        ['V-108', 'LogiTrans Corp', 'Naomi Tanaka', 'n.tanaka@logitrans.com', '555-0108', '2026-04-30'],
        ['V-109', 'FreshAir HVAC', 'Derek Stone', 'd.stone@freshair.com', '555-0109', '2025-10-15'],
        ['V-110', 'BuildRight Maint', 'Sofia Andersson', 's.andersson@buildright.se', '555-0110', '2026-02-28'],
    ]
    for r, row_data in enumerate(vendor_data, 2):
        for c, val in enumerate(row_data, 1):
            ws7.cell(row=r, column=c, value=val)

    # --- Sheet8: Training Schedule ---
    ws8 = wb.create_sheet('Sheet8')
    headers8 = ['Course', 'Instructor', 'Department', 'Date', 'Duration (hrs)', 'Capacity']
    for c, h in enumerate(headers8, 1):
        ws8.cell(row=1, column=c, value=h)
        ws8.cell(row=1, column=c).font = Font(bold=True)
    train_data = [
        ['Python Advanced', 'Dr. Alan Turing', 'Engineering', '2025-04-10', 8, 30],
        ['Leadership 101', 'Maria Gonzalez', 'All', '2025-04-15', 4, 50],
        ['Data Analytics', 'Raj Patel', 'Finance', '2025-04-20', 6, 25],
        ['Cybersecurity Basics', 'Kim Nguyen', 'All', '2025-05-01', 8, 40],
        ['Project Management', 'Thomas Mueller', 'Product', '2025-05-05', 6, 35],
        ['Cloud Architecture', 'Lisa Wang', 'Engineering', '2025-05-12', 8, 20],
        ['Sales Negotiation', 'Chris O\'Brien', 'Sales', '2025-05-18', 4, 30],
        ['UX Design Thinking', 'Yuki Sato', 'Product', '2025-05-25', 6, 25],
        ['Compliance Training', 'Anna Schmidt', 'Legal', '2025-06-01', 4, 60],
        ['Agile Methodology', 'Omar Hassan', 'Engineering', '2025-06-08', 8, 35],
    ]
    for r, row_data in enumerate(train_data, 2):
        for c, val in enumerate(row_data, 1):
            ws8.cell(row=r, column=c, value=val)

    # --- Sheet9: KPI Dashboard ---
    ws9 = wb.create_sheet('Sheet9')
    headers9 = ['Metric', 'Target', 'Actual', 'Status', 'Trend', 'Owner']
    for c, h in enumerate(headers9, 1):
        ws9.cell(row=1, column=c, value=h)
        ws9.cell(row=1, column=c).font = Font(bold=True)
    kpi_data = [
        ['Revenue Growth', '15%', '17.2%', 'On Track', 'Up', 'CFO'],
        ['Customer Retention', '90%', '88.5%', 'At Risk', 'Down', 'VP Sales'],
        ['NPS Score', '72', '75', 'On Track', 'Up', 'VP Product'],
        ['Employee Satisfaction', '4.2', '4.0', 'At Risk', 'Stable', 'CHRO'],
        ['Time to Market', '90 days', '85 days', 'On Track', 'Up', 'VP Engineering'],
        ['Cost per Acquisition', '$45', '$42', 'On Track', 'Down', 'VP Marketing'],
        ['Uptime SLA', '99.9%', '99.95%', 'On Track', 'Up', 'VP Ops'],
        ['Bug Resolution Time', '48 hrs', '52 hrs', 'At Risk', 'Up', 'VP Engineering'],
        ['Support Ticket Volume', '500/mo', '478/mo', 'On Track', 'Down', 'VP Support'],
        ['Training Completion', '85%', '91%', 'On Track', 'Up', 'CHRO'],
    ]
    for r, row_data in enumerate(kpi_data, 2):
        for c, val in enumerate(row_data, 1):
            ws9.cell(row=r, column=c, value=val)

    # --- Sheet10: Meeting Notes ---
    ws10 = wb.create_sheet('Sheet10')
    headers10 = ['Date', 'Meeting', 'Attendees', 'Key Decision', 'Action Item', 'Owner']
    for c, h in enumerate(headers10, 1):
        ws10.cell(row=1, column=c, value=h)
        ws10.cell(row=1, column=c).font = Font(bold=True)
    meeting_data = [
        ['2025-03-03', 'Sprint Planning', 'Engineering Team', 'Prioritize auth refactor', 'Create Jira tickets', 'Priya Sharma'],
        ['2025-03-05', 'Budget Review', 'Finance + Dept Heads', 'Cut travel 10%', 'Update forecasts', 'David Kim'],
        ['2025-03-07', 'Product Roadmap', 'Product + Engineering', 'Delay Feature X to Q3', 'Update timeline', 'Fatima Al-Hassan'],
        ['2025-03-10', 'All Hands', 'All Staff', 'Announce Q2 goals', 'Distribute goals doc', 'CEO'],
        ['2025-03-12', 'Design Review', 'Product + UX', 'Approve new dashboard', 'Start implementation', 'Yuki Sato'],
        ['2025-03-14', 'Sales Forecast', 'Sales + Finance', 'Revise Q2 targets up', 'Update CRM pipeline', 'Emma Thompson'],
        ['2025-03-17', 'Security Audit', 'IT + Legal', 'Fix 3 critical vulns', 'Patch by March 25', 'Carlos Rivera'],
        ['2025-03-19', 'Vendor Negotiation', 'Procurement + Legal', 'Extend CloudNet contract', 'Draft amendment', 'Linda Park'],
        ['2025-03-21', 'Sprint Retro', 'Engineering Team', 'Improve code reviews', 'Update PR template', 'Wei Zhang'],
        ['2025-03-24', 'Marketing Campaign', 'Marketing + Sales', 'Launch email series', 'Create content calendar', 'Marcus Johnson'],
    ]
    for r, row_data in enumerate(meeting_data, 2):
        for c, val in enumerate(row_data, 1):
            ws10.cell(row=r, column=c, value=val)

    # NO tab colors set - that is the task for the agent
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
