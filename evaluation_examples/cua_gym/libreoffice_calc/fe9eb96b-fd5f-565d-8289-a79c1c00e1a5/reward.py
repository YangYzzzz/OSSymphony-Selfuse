"""
Reward Script: HR Salary Band Lookup with VLOOKUP and Band Status
Task ID: calc_hr_salary_band_lookup_009
Domain: libreoffice_calc
Scoring:
  Precondition gate: Pay Grades sheet intact (if not, return 0.0)
  Component 1 (0.45): E2:E78 contain correct VLOOKUP formulas for salary band midpoint
  Component 2 (0.35): F2:F78 contain correct IF formulas for band status (Below/Within/Above)
  Component 3 (0.20): E2:E78 formatted as currency $#,##0
Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_salary_band_lookup_009'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify required sheets exist (precondition gate)
    if 'Employees' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Employees' not found")
        print("REWARD: 0.0")
        return 0.0

    if 'Pay Grades' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Pay Grades' not found")
        print("REWARD: 0.0")
        return 0.0

    ws_emp = wb['Employees']
    ws_pg = wb['Pay Grades']

    # -----------------------------------------------------------------------
    # Precondition Gate: Pay Grades sheet must have correct data (not modified)
    # This is a constraint gate, NOT a scoring component — it does not award points.
    # If Pay Grades is corrupted, we cannot meaningfully evaluate VLOOKUP correctness.
    # -----------------------------------------------------------------------
    try:
        expected_grades = {
            'G1': 45000, 'G2': 58000, 'G3': 75000, 'G4': 95000, 'G5': 118000
        }
        pay_grades_issues = []
        for row, (grade, mid) in enumerate(expected_grades.items(), 2):
            a_val = ws_pg.cell(row=row, column=1).value
            c_val = ws_pg.cell(row=row, column=3).value
            if str(a_val) != grade or c_val != mid:
                pay_grades_issues.append(
                    f"row {row}: grade={repr(a_val)}, mid={repr(c_val)}"
                )
                print(f"WARNING: Pay Grades data corrupted at row {row}: grade={repr(a_val)}, mid={repr(c_val)}")

        if len(pay_grades_issues) > 0:
            print("CRITICAL: Pay Grades sheet is corrupted — cannot proceed with VLOOKUP verification")
            print("REWARD: 0.0")
            return 0.0
        else:
            print("INFO: Pay Grades table is intact (precondition satisfied)")
    except Exception as e:
        print(f"WARNING: Could not verify Pay Grades integrity — {e}")

    # -----------------------------------------------------------------------
    # Component 1: E2:E78 contain VLOOKUP formulas for salary band midpoint (0.45 points)
    # Expected pattern: =VLOOKUP(C<row>,'Pay Grades'.$A:$D,3,0)
    # Key requirements:
    #   - Must reference C<row> (pay grade column)
    #   - Must reference 'Pay Grades' sheet
    #   - Must return column 3 (Midpoint)
    #   - Must be exact match (last arg = 0 or FALSE)
    # -----------------------------------------------------------------------
    try:
        vlookup_count = 0
        vlookup_total = 77  # rows 2-78
        vlookup_errors = []

        for row in range(2, 79):
            cell = ws_emp.cell(row=row, column=5)  # Column E
            val = cell.value
            if val is None:
                vlookup_errors.append(f"E{row}: empty")
            elif not isinstance(val, str):
                vlookup_errors.append(f"E{row}: not a formula string: {repr(val)}")
            else:
                normalized = val.strip()
                # Flexible check: must be a VLOOKUP, reference 'Pay Grades', col 3, exact match (0 or FALSE)
                if (normalized.upper().startswith('=VLOOKUP') and
                        'PAY GRADES' in normalized.upper() and
                        re.search(r',\s*3\s*,', normalized) and
                        re.search(r',\s*(0|FALSE)\s*\)', normalized, re.IGNORECASE)):
                    vlookup_count += 1
                else:
                    vlookup_errors.append(f"E{row}: incorrect formula: {repr(normalized[:80])}")

        if vlookup_count == vlookup_total:
            print(f"PASS: Component 1 — All {vlookup_total} VLOOKUP formulas in E2:E78 are correct (0.45 pts)")
            total_score += 0.45
        elif vlookup_count >= vlookup_total * 0.9:
            # Partial credit for near-complete implementation (>=90% correct)
            print(f"PARTIAL: Component 1 — {vlookup_count}/{vlookup_total} VLOOKUP formulas correct (0.22 pts)")
            if vlookup_errors:
                print(f"  First errors: {vlookup_errors[:3]}")
            total_score += 0.22
        else:
            print(f"FAIL: Component 1 — Only {vlookup_count}/{vlookup_total} VLOOKUP formulas correct")
            if vlookup_errors:
                print(f"  First errors: {vlookup_errors[:3]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: F2:F78 contain IF formulas for band status (0.35 points)
    # Expected pattern: =IF(D<row><E<row>*0.9,"Below Band",IF(D<row>>E<row>*1.1,"Above Band","Within Band"))
    # Key requirements:
    #   - Must reference D<row> and E<row>
    #   - Must use 0.9 threshold for "Below Band"
    #   - Must use 1.1 threshold for "Above Band"
    #   - Must produce exactly: "Below Band", "Within Band", "Above Band"
    # -----------------------------------------------------------------------
    try:
        if_count = 0
        if_total = 77  # rows 2-78
        if_errors = []

        for row in range(2, 79):
            cell = ws_emp.cell(row=row, column=6)  # Column F
            val = cell.value
            if val is None:
                if_errors.append(f"F{row}: empty")
            elif not isinstance(val, str):
                if_errors.append(f"F{row}: not a formula string: {repr(val)}")
            else:
                normalized_upper = val.strip().upper().replace(' ', '')
                # Check for IF formula with correct structure
                has_if = normalized_upper.startswith('=IF(')
                # Check for band labels (with or without spaces in the formula)
                has_below_band = 'BELOWBAND' in normalized_upper
                has_above_band = 'ABOVEBAND' in normalized_upper
                has_within_band = 'WITHINBAND' in normalized_upper
                has_09_threshold = '0.9' in normalized_upper
                has_11_threshold = '1.1' in normalized_upper
                has_d_ref = f'D{row}' in val
                has_e_ref = f'E{row}' in val

                if (has_if and has_below_band and has_above_band and has_within_band and
                        has_09_threshold and has_11_threshold and has_d_ref and has_e_ref):
                    if_count += 1
                else:
                    if_errors.append(
                        f"F{row}: incorrect formula: {repr(val[:80])}; "
                        f"has_if={has_if}, below={has_below_band}, above={has_above_band}, "
                        f"within={has_within_band}, 0.9={has_09_threshold}, 1.1={has_11_threshold}"
                    )

        if if_count == if_total:
            print(f"PASS: Component 2 — All {if_total} IF band-status formulas in F2:F78 are correct (0.35 pts)")
            total_score += 0.35
        elif if_count >= if_total * 0.9:
            print(f"PARTIAL: Component 2 — {if_count}/{if_total} IF formulas correct (0.17 pts)")
            if if_errors:
                print(f"  First errors: {if_errors[:3]}")
            total_score += 0.17
        else:
            print(f"FAIL: Component 2 — Only {if_count}/{if_total} IF formulas correct")
            if if_errors:
                print(f"  First errors: {if_errors[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: E2:E78 formatted as currency $#,##0 (0.20 points)
    # Task spec says: "E2:E78 formatted as currency $#,##0"
    # The initial file has E2:E78 empty with 'General' format,
    # so any currency format in E2:E78 reflects the task action.
    # -----------------------------------------------------------------------
    try:
        currency_count = 0
        currency_total = 77  # rows 2-78
        currency_errors = []

        for row in range(2, 79):
            cell = ws_emp.cell(row=row, column=5)  # Column E
            fmt = cell.number_format or 'General'
            # Accept '$#,##0' and similar currency variants
            if fmt != 'General' and (fmt.startswith('$#,##0') or '$' in fmt):
                currency_count += 1
            else:
                currency_errors.append(f"E{row}: format={repr(fmt)}")

        if currency_count == currency_total:
            print(f"PASS: Component 3 — All {currency_total} cells in E2:E78 have currency format (0.20 pts)")
            total_score += 0.20
        elif currency_count >= currency_total * 0.9:
            print(f"PARTIAL: Component 3 — {currency_count}/{currency_total} cells have currency format (0.10 pts)")
            if currency_errors:
                print(f"  First errors: {currency_errors[:3]}")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — Only {currency_count}/{currency_total} cells have currency format '$#,##0'")
            if currency_errors:
                print(f"  First errors: {currency_errors[:3]}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
