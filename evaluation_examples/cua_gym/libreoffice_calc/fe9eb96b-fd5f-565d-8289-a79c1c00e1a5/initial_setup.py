"""
Initial Setup: HR Salary Band Lookup
Task ID: calc_hr_salary_band_lookup_009
Domain: libreoffice_calc

Creates an Employees sheet with 77 employees (rows 2-78) and a Pay Grades sheet.
Columns E (Band Midpoint) and F (Band Status) are intentionally LEFT EMPTY.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_salary_band_lookup_009'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # -------------------------------------------------------------------------
    # Sheet 1: Employees
    # -------------------------------------------------------------------------
    ws_emp = wb.active
    ws_emp.title = 'Employees'

    # Headers (row 1)
    emp_headers = ['Emp ID', 'Name', 'Pay Grade', 'Current Salary', 'Band Midpoint', 'Band Status']
    for col, h in enumerate(emp_headers, 1):
        cell = ws_emp.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Employee data — realistic names, pay grades G1-G5, salary values
    # Columns: Emp ID, Name, Pay Grade, Current Salary
    # E and F are LEFT EMPTY intentionally
    employee_data = [
        ('E001', 'Sarah Chen',          'G3', 72500),
        ('E002', 'Marcus Johnson',      'G2', 49800),
        ('E003', 'Priya Patel',         'G4', 91000),
        ('E004', 'Derek Williams',      'G1', 38500),
        ('E005', 'Laura Nguyen',        'G3', 68000),
        ('E006', 'James O\'Brien',      'G5', 115000),
        ('E007', 'Amara Okafor',        'G2', 53200),
        ('E008', 'Thomas Schulz',       'G4', 87500),
        ('E009', 'Mei-Ling Zhang',      'G3', 75000),
        ('E010', 'Carlos Rivera',       'G1', 41200),
        ('E011', 'Fatima Al-Hassan',    'G5', 122000),
        ('E012', 'Nathan Brooks',       'G2', 47000),
        ('E013', 'Isabelle Dupont',     'G3', 69500),
        ('E014', 'Kwame Asante',        'G4', 93500),
        ('E015', 'Rachel Kim',          'G1', 37800),
        ('E016', 'Oliver Petrov',       'G5', 118500),
        ('E017', 'Dana Foster',         'G2', 51000),
        ('E018', 'Yuki Tanaka',         'G3', 73000),
        ('E019', 'Brendan Murphy',      'G4', 89000),
        ('E020', 'Alicia Morales',      'G1', 43500),
        ('E021', 'Sven Larsson',        'G5', 110000),
        ('E022', 'Tanya Robinson',      'G2', 54500),
        ('E023', 'Haruto Yamamoto',     'G3', 66000),
        ('E024', 'Grace Adeyemi',       'G4', 95000),
        ('E025', 'Patrick Sullivan',    'G1', 36500),
        ('E026', 'Nadia Kozlov',        'G5', 125000),
        ('E027', 'Michael Thompson',    'G2', 48000),
        ('E028', 'Elisa Fontaine',      'G3', 77000),
        ('E029', 'David Chukwu',        'G4', 86000),
        ('E030', 'Anna Sorensen',       'G1', 40500),
        ('E031', 'Roberto Martinez',    'G5', 117500),
        ('E032', 'Sandra Lee',          'G2', 52500),
        ('E033', 'Arjun Sharma',        'G3', 70000),
        ('E034', 'Claudia Weber',       'G4', 92000),
        ('E035', 'Jack Harrison',       'G1', 39000),
        ('E036', 'Fatou Diallo',        'G5', 120000),
        ('E037', 'Liam Conway',         'G2', 46500),
        ('E038', 'Akemi Nakamura',      'G3', 74500),
        ('E039', 'Victor Osei',         'G4', 88500),
        ('E040', 'Maria Santos',        'G1', 42000),
        ('E041', 'Ethan Clarke',        'G5', 113000),
        ('E042', 'Ingrid Bakke',        'G2', 55000),
        ('E043', 'Samuel Abiodun',      'G3', 67500),
        ('E044', 'Hana Kimura',         'G4', 94000),
        ('E045', 'Kevin Walsh',         'G1', 37200),
        ('E046', 'Mihaela Ionescu',     'G5', 119000),
        ('E047', 'Ben Watkins',         'G2', 50500),
        ('E048', 'Zara Ahmed',          'G3', 71500),
        ('E049', 'Cormac Flynn',        'G4', 90000),
        ('E050', 'Lucia Fernandez',     'G1', 44000),
        ('E051', 'Dmitri Volkov',       'G5', 116000),
        ('E052', 'Sophie Dupuis',       'G2', 47500),
        ('E053', 'Emmanuel Nwosu',      'G3', 76000),
        ('E054', 'Rin Watanabe',        'G4', 85500),
        ('E055', 'Tyler Henderson',     'G1', 38800),
        ('E056', 'Adaeze Obi',          'G5', 123000),
        ('E057', 'Mikael Eriksson',     'G2', 53700),
        ('E058', 'Camille Bonnet',      'G3', 69000),
        ('E059', 'Ahmad Khalid',        'G4', 96500),
        ('E060', 'Patricia Brown',      'G1', 41800),
        ('E061', 'Leon Fischer',        'G5', 111500),
        ('E062', 'Chioma Okafor',       'G2', 51500),
        ('E063', 'Ryo Suzuki',          'G3', 72000),
        ('E064', 'Hannah McGrath',      'G4', 87000),
        ('E065', 'Babatunde Adeyemi',   'G1', 36200),
        ('E066', 'Elena Morozova',      'G5', 121000),
        ('E067', 'Finn Johansson',      'G2', 49200),
        ('E068', 'Layla Ibrahim',       'G3', 75500),
        ('E069', 'Stephen Okonkwo',     'G4', 91500),
        ('E070', 'Yun Li',              'G1', 43200),
        ('E071', 'Anastasia Petrov',    'G5', 114000),
        ('E072', 'Declan Moore',        'G2', 56000),
        ('E073', 'Ngozi Eze',           'G3', 68500),
        ('E074', 'Takuya Morita',       'G4', 93000),
        ('E075', 'Brigitte Hoffman',    'G1', 40000),
        ('E076', 'Kofi Mensah',         'G5', 126000),
        ('E077', 'Svetlana Ivanova',    'G2', 48500),
    ]

    for row_idx, (emp_id, name, grade, salary) in enumerate(employee_data, 2):
        ws_emp.cell(row=row_idx, column=1, value=emp_id)
        ws_emp.cell(row=row_idx, column=2, value=name)
        ws_emp.cell(row=row_idx, column=3, value=grade)
        ws_emp.cell(row=row_idx, column=4, value=salary)
        # Columns E (5) and F (6) intentionally left empty

    # Column widths
    ws_emp.column_dimensions['A'].width = 10
    ws_emp.column_dimensions['B'].width = 24
    ws_emp.column_dimensions['C'].width = 12
    ws_emp.column_dimensions['D'].width = 16
    ws_emp.column_dimensions['E'].width = 16
    ws_emp.column_dimensions['F'].width = 14

    # -------------------------------------------------------------------------
    # Sheet 2: Pay Grades
    # -------------------------------------------------------------------------
    ws_pg = wb.create_sheet('Pay Grades')

    pg_headers = ['Pay Grade', 'Min', 'Midpoint', 'Max']
    for col, h in enumerate(pg_headers, 1):
        cell = ws_pg.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    pay_grade_data = [
        ('G1', 32000, 45000, 58000),
        ('G2', 42000, 58000, 74000),
        ('G3', 58000, 75000, 92000),
        ('G4', 75000, 95000, 115000),
        ('G5', 95000, 118000, 141000),
    ]

    for row_idx, (grade, min_sal, mid_sal, max_sal) in enumerate(pay_grade_data, 2):
        ws_pg.cell(row=row_idx, column=1, value=grade)
        ws_pg.cell(row=row_idx, column=2, value=min_sal)
        ws_pg.cell(row=row_idx, column=3, value=mid_sal)
        ws_pg.cell(row=row_idx, column=4, value=max_sal)

    ws_pg.column_dimensions['A'].width = 12
    ws_pg.column_dimensions['B'].width = 12
    ws_pg.column_dimensions['C'].width = 12
    ws_pg.column_dimensions['D'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Employees sheet: {len(employee_data)} rows (rows 2-78)')
    print(f'  Pay Grades sheet: {len(pay_grade_data)} rows (rows 2-6)')
    print(f'  Columns E and F in Employees are EMPTY (no formulas)')


create_initial()
