"""
Initial Setup: Create performance report spreadsheet with monthly revenue and growth rate data
Task ID: calc_gg2_013
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_gg2_013'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Performance Sheet ---
    ws = wb.active
    ws.title = 'Performance'

    # Headers
    headers = ['Month', 'Revenue', 'Growth Rate']
    header_font = Font(name='Calibri', size=11, bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    # Monthly data - realistic revenue in millions and growth rate percentages
    months = [
        'January', 'February', 'March', 'April', 'May', 'June',
        'July', 'August', 'September', 'October', 'November', 'December'
    ]
    revenues = [
        2.45, 2.68, 3.12, 3.05, 3.47, 3.89,
        3.72, 4.15, 4.38, 4.62, 5.01, 5.34
    ]
    growth_rates = [
        0.032, 0.094, 0.164, -0.022, 0.138, 0.121,
        -0.044, 0.116, 0.055, 0.055, 0.084, 0.066
    ]

    for r, (month, rev, gr) in enumerate(zip(months, revenues, growth_rates), 2):
        ws.cell(row=r, column=1, value=month)
        ws.cell(row=r, column=2, value=rev)
        cell_gr = ws.cell(row=r, column=3, value=gr)
        cell_gr.number_format = '0.0%'

    # Format revenue column as number with 2 decimals
    for r in range(2, 14):
        ws.cell(row=r, column=2).number_format = '#,##0.00'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14

    # NO charts - the task is to create the combination chart
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
