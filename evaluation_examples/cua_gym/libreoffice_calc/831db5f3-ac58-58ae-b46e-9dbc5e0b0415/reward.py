"""
Reward Script: Inventory Valuation with FIFO, Market Value, LCM Adjustment
Task ID: calc_fin_inventory_valuation_058
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Headers F1='FIFO Value', G1='Market Value', H1='LCM Adjustment' — 0.15 pts
  Component 2: F2:F60 formulas =C{r}*D{r} — 0.20 pts
  Component 3: G2:G60 formulas =C{r}*E{r} — 0.15 pts
  Component 4: H2:H60 formulas =MIN(F{r},G{r})-F{r} — 0.15 pts
  Component 5: Row 61 SUM totals, bold, currency format — 0.10 pts
  Component 6: F2:H60 currency formatted — 0.05 pts
  Component 7: Conditional formatting H2:H60 value < -500 red bg — 0.10 pts
  Component 8: Rows sorted by FIFO value descending — 0.10 pts
  Total: 1.00
"""

import os
import openpyxl
from openpyxl.utils import get_column_letter

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_inventory_valuation_058'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Ensure the Inventory sheet exists
    if 'Inventory' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Inventory' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Inventory']

    # -------------------------------------------------------------------
    # Component 1: Headers in row 1 — F1='FIFO Value', G1='Market Value',
    #              H1='LCM Adjustment' (0.15 points)
    # These columns are EMPTY in the initial file, so this tests task completion.
    # -------------------------------------------------------------------
    try:
        f1 = ws.cell(row=1, column=6).value
        g1 = ws.cell(row=1, column=7).value
        h1 = ws.cell(row=1, column=8).value

        expected_headers = {
            6: 'FIFO Value',
            7: 'Market Value',
            8: 'LCM Adjustment'
        }

        headers_ok = (
            str(f1).strip() == 'FIFO Value' and
            str(g1).strip() == 'Market Value' and
            str(h1).strip() == 'LCM Adjustment'
        )

        if headers_ok:
            print(f"PASS: Component 1 — Headers correct: F1='{f1}', G1='{g1}', H1='{h1}' (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — Headers incorrect: F1='{f1}', G1='{g1}', H1='{h1}'")
            print(f"  Expected: F1='FIFO Value', G1='Market Value', H1='LCM Adjustment'")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------
    # Component 2: F2:F60 formulas =C{r}*D{r} (0.20 points)
    # Initial file has F2:F60 all empty, so any formula presence is task-completion.
    # Check at least 50 of 59 rows have the correct =C{r}*D{r} formula pattern.
    # -------------------------------------------------------------------
    try:
        correct_f_formulas = 0
        total_f_rows = 59  # rows 2-60

        for row in range(2, 61):
            cell_val = ws.cell(row=row, column=6).value
            if cell_val is not None:
                formula_str = str(cell_val).strip().upper().replace(' ', '')
                expected = f'=C{row}*D{row}'.upper()
                if formula_str == expected:
                    correct_f_formulas += 1

        pct_correct = correct_f_formulas / total_f_rows
        if pct_correct >= 0.85:
            print(f"PASS: Component 2 — F column formulas: {correct_f_formulas}/{total_f_rows} rows have =C{{r}}*D{{r}} (0.20 pts)")
            total_score += 0.20
        elif pct_correct >= 0.50:
            partial = 0.10
            print(f"PARTIAL: Component 2 — F column formulas: {correct_f_formulas}/{total_f_rows} rows correct (0.10 pts partial)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — F column formulas: only {correct_f_formulas}/{total_f_rows} rows have =C{{r}}*D{{r}}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------
    # Component 3: G2:G60 formulas =C{r}*E{r} (0.15 points)
    # Initial file has G2:G60 all empty.
    # -------------------------------------------------------------------
    try:
        correct_g_formulas = 0
        total_g_rows = 59  # rows 2-60

        for row in range(2, 61):
            cell_val = ws.cell(row=row, column=7).value
            if cell_val is not None:
                formula_str = str(cell_val).strip().upper().replace(' ', '')
                expected = f'=C{row}*E{row}'.upper()
                if formula_str == expected:
                    correct_g_formulas += 1

        pct_correct_g = correct_g_formulas / total_g_rows
        if pct_correct_g >= 0.85:
            print(f"PASS: Component 3 — G column formulas: {correct_g_formulas}/{total_g_rows} rows have =C{{r}}*E{{r}} (0.15 pts)")
            total_score += 0.15
        elif pct_correct_g >= 0.50:
            partial = 0.07
            print(f"PARTIAL: Component 3 — G column formulas: {correct_g_formulas}/{total_g_rows} rows correct (0.07 pts partial)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — G column formulas: only {correct_g_formulas}/{total_g_rows} rows have =C{{r}}*E{{r}}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------
    # Component 4: H2:H60 formulas =MIN(F{r},G{r})-F{r} (0.15 points)
    # Initial file has H2:H60 all empty.
    # -------------------------------------------------------------------
    try:
        correct_h_formulas = 0
        total_h_rows = 59  # rows 2-60

        for row in range(2, 61):
            cell_val = ws.cell(row=row, column=8).value
            if cell_val is not None:
                formula_str = str(cell_val).strip().upper().replace(' ', '')
                expected = f'=MIN(F{row},G{row})-F{row}'.upper()
                if formula_str == expected:
                    correct_h_formulas += 1

        pct_correct_h = correct_h_formulas / total_h_rows
        if pct_correct_h >= 0.85:
            print(f"PASS: Component 4 — H column formulas: {correct_h_formulas}/{total_h_rows} rows have =MIN(F{{r}},G{{r}})-F{{r}} (0.15 pts)")
            total_score += 0.15
        elif pct_correct_h >= 0.50:
            partial = 0.07
            print(f"PARTIAL: Component 4 — H column formulas: {correct_h_formulas}/{total_h_rows} rows correct (0.07 pts partial)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — H column formulas: only {correct_h_formulas}/{total_h_rows} rows have =MIN(F{{r}},G{{r}})-F{{r}}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -------------------------------------------------------------------
    # Component 5: Row 61 SUM totals with bold formatting and currency format (0.10 points)
    # Initial file has no row 61 data; golden adds SUM formulas + bold + currency.
    # -------------------------------------------------------------------
    try:
        f61 = ws.cell(row=61, column=6).value
        g61 = ws.cell(row=61, column=7).value
        h61 = ws.cell(row=61, column=8).value
        f61_bold = ws.cell(row=61, column=6).font.bold
        g61_bold = ws.cell(row=61, column=7).font.bold
        h61_bold = ws.cell(row=61, column=8).font.bold
        f61_fmt = ws.cell(row=61, column=6).number_format
        g61_fmt = ws.cell(row=61, column=7).number_format
        h61_fmt = ws.cell(row=61, column=8).number_format

        # Check formulas
        f61_formula_ok = f61 is not None and 'SUM' in str(f61).upper() and 'F2' in str(f61).upper()
        g61_formula_ok = g61 is not None and 'SUM' in str(g61).upper() and 'G2' in str(g61).upper()
        h61_formula_ok = h61 is not None and 'SUM' in str(h61).upper() and 'H2' in str(h61).upper()

        # Check bold
        totals_bold = bool(f61_bold) and bool(g61_bold) and bool(h61_bold)

        # Check currency format (should contain '$' or '#,##0')
        def is_currency_fmt(fmt):
            if fmt is None:
                return False
            return '$' in str(fmt) or ('#,##0' in str(fmt))

        totals_currency = is_currency_fmt(f61_fmt) and is_currency_fmt(g61_fmt) and is_currency_fmt(h61_fmt)

        if f61_formula_ok and g61_formula_ok and h61_formula_ok and totals_bold and totals_currency:
            print(f"PASS: Component 5 — Row 61 totals: SUM formulas + bold + currency format (0.10 pts)")
            total_score += 0.10
        elif f61_formula_ok and g61_formula_ok and h61_formula_ok:
            print(f"PARTIAL: Component 5 — Row 61 SUM formulas present but bold={totals_bold}, currency={totals_currency} (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — Row 61: F61={f61}, G61={g61}, H61={h61}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # -------------------------------------------------------------------
    # Component 6: F2:H60 currency formatted (0.05 points)
    # Check a sample of cells in columns F, G, H rows 2-60 have currency format.
    # -------------------------------------------------------------------
    try:
        currency_check_rows = [2, 5, 15, 30, 45, 60]
        currency_count = 0
        total_checks = len(currency_check_rows) * 3  # 3 columns

        for row in currency_check_rows:
            for col in [6, 7, 8]:
                fmt = ws.cell(row=row, column=col).number_format
                if fmt and ('$' in str(fmt) or '#,##0' in str(fmt)):
                    currency_count += 1

        if currency_count >= total_checks * 0.8:
            print(f"PASS: Component 6 — Currency format: {currency_count}/{total_checks} sampled cells have currency format (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 — Currency format: only {currency_count}/{total_checks} sampled cells have currency format")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # -------------------------------------------------------------------
    # Component 7: Conditional formatting on H2:H60, value < -500 => red background (0.10 points)
    # Initial file has no CF rules; golden adds them.
    # -------------------------------------------------------------------
    try:
        cf_rules = ws.conditional_formatting
        found_cf = False

        for cf_range in cf_rules._cf_rules:
            cf_range_str = str(cf_range)
            # Look for a CF rule on H column
            if 'H' in cf_range_str:
                for rule in cf_rules._cf_rules[cf_range]:
                    rule_type = getattr(rule, 'type', None)
                    op = getattr(rule, 'operator', None)
                    formula = getattr(rule, 'formula', None)

                    # Check it's a cellIs lessThan -500 rule
                    is_less_than_500 = False
                    if rule_type == 'cellIs' and op == 'lessThan':
                        if formula and len(formula) > 0:
                            try:
                                threshold = float(str(formula[0]).strip())
                                is_less_than_500 = threshold <= -500
                            except (ValueError, TypeError):
                                pass

                    # Check it has red fill
                    has_red_fill = False
                    if hasattr(rule, 'dxf') and rule.dxf is not None:
                        dxf = rule.dxf
                        if dxf.fill is not None:
                            try:
                                fg_rgb = dxf.fill.fgColor.rgb
                                # Red: contains FF0000 somewhere in the ARGB
                                if fg_rgb and ('FF0000' in fg_rgb.upper() or fg_rgb.upper() == 'FFFF0000'):
                                    has_red_fill = True
                            except Exception:
                                pass

                    if is_less_than_500 and has_red_fill:
                        found_cf = True
                        print(f"PASS: Component 7 — CF found on H column: type={rule_type}, op={op}, "
                              f"threshold={formula}, red_fill=True (0.10 pts)")
                        break
                if found_cf:
                    break

        if found_cf:
            total_score += 0.10
        else:
            print(f"FAIL: Component 7 — No conditional formatting found on H column with lessThan -500 + red background")
            # Show what CF rules exist for debugging
            for cf_range in cf_rules._cf_rules:
                print(f"  Found CF range: {cf_range}")
                for rule in cf_rules._cf_rules[cf_range]:
                    print(f"    rule: type={getattr(rule, 'type', None)}, op={getattr(rule, 'operator', None)}, formula={getattr(rule, 'formula', None)}")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    # -------------------------------------------------------------------
    # Component 8: Rows sorted by FIFO value descending (0.10 points)
    # The golden file rows 2-60 are reordered vs initial. We verify the sort
    # by computing FIFO values (units * fifo_cost) for each row and checking
    # they are in non-increasing order.
    # -------------------------------------------------------------------
    try:
        fifo_values = []
        for row in range(2, 61):
            units = ws.cell(row=row, column=3).value
            fifo_cost = ws.cell(row=row, column=4).value
            if units is not None and fifo_cost is not None:
                try:
                    fifo_val = float(units) * float(fifo_cost)
                    fifo_values.append(fifo_val)
                except (ValueError, TypeError):
                    fifo_values.append(None)

        # Check that values are sorted in non-increasing order
        valid_vals = [v for v in fifo_values if v is not None]
        if len(valid_vals) >= 50:
            is_sorted_desc = all(valid_vals[i] >= valid_vals[i+1] for i in range(len(valid_vals)-1))
            if is_sorted_desc:
                print(f"PASS: Component 8 — Sort: {len(valid_vals)} rows sorted by FIFO value descending (0.10 pts)")
                total_score += 0.10
            else:
                # Count violations
                violations = sum(1 for i in range(len(valid_vals)-1) if valid_vals[i] < valid_vals[i+1])
                print(f"FAIL: Component 8 — Sort order incorrect: {violations} ordering violations in {len(valid_vals)} rows")
        else:
            print(f"FAIL: Component 8 — Not enough data rows to verify sort: {len(valid_vals)} valid rows")
    except Exception as e:
        print(f"ERROR: Component 8 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
