"""
Initial Setup: Inventory valuation spreadsheet (pre-task state)
Task ID: calc_fin_inventory_valuation_058
Domain: libreoffice_calc

Creates an Inventory sheet with SKU, Description, Units, FIFO Cost, Replacement Cost
(columns A-E). Columns F, G, H are intentionally left empty — those will be filled
by the agent during the task.
"""

import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_inventory_valuation_058'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventory'

    # --- Row 1: Headers (A-E only; F, G, H must be empty) ---
    headers = ['SKU', 'Description', 'Units', 'FIFO Cost', 'Replacement Cost']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        # Light grey header styling (not bold — boldness will be task outcome)
        cell.font = Font(name='Calibri', size=11)

    # --- Rows 2-60: Inventory data (59 items) ---
    # Realistic warehouse / retail inventory items
    # FIFO Cost = original purchase cost per unit
    # Replacement Cost = current market price per unit
    inventory_items = [
        # SKU,         Description,                               Units, FIFO Cost, Replacement Cost
        ('SKU-1001', 'Wireless Bluetooth Headphones Model X200',   145,   42.50,   38.75),
        ('SKU-1002', 'USB-C Charging Cable 6ft 3-Pack',            820,    4.20,    3.95),
        ('SKU-1003', 'Ergonomic Office Chair - Mesh Back',          38,  185.00,  210.00),
        ('SKU-1004', 'Standing Desk Converter 36-inch',             22,  129.99,  115.00),
        ('SKU-1005', 'Mechanical Keyboard RGB Backlit',             97,   65.00,   58.50),
        ('SKU-1006', 'Laptop Cooling Pad with 5 Fans',            153,   22.75,   19.90),
        ('SKU-1007', 'Dual Monitor Arm Adjustable',                64,   48.00,   52.00),
        ('SKU-1008', 'HD Webcam 1080p with Ring Light',           211,   34.50,   29.00),
        ('SKU-1009', 'Noise-Cancelling Earbud Pro Series',        178,   55.00,   47.25),
        ('SKU-1010', 'Portable SSD 1TB USB 3.2',                   89,   78.00,   71.50),
        ('SKU-1011', 'Wireless Mouse Ergonomic Vertical',         302,   18.40,   16.80),
        ('SKU-1012', 'External DVD/Blu-ray Drive USB',             47,   32.00,   28.50),
        ('SKU-1013', 'Smart Surge Protector 8-Outlet Wi-Fi',      136,   29.95,   27.00),
        ('SKU-1014', 'Cable Management Box Organiser',            445,    8.50,    7.75),
        ('SKU-1015', 'Monitor Privacy Screen 27-inch',             58,   22.00,   20.50),
        ('SKU-1016', 'Graphics Tablet Drawing Pad Medium',         31,   95.00,   88.00),
        ('SKU-1017', 'Label Printer Thermal Desktop',              19,  112.00,  118.00),
        ('SKU-1018', 'Laptop Backpack Anti-Theft 15.6in',         263,   35.00,   31.50),
        ('SKU-1019', 'LED Desk Lamp with USB Charging Port',      192,   24.80,   22.00),
        ('SKU-1020', 'Laser Printer Toner Cartridge Black XL',     74,   42.00,   38.00),
        ('SKU-1021', 'Multifunction Printer Ink Set CMYK',         88,   28.50,   26.00),
        ('SKU-1022', 'Adjustable Laptop Stand Aluminium',         172,   19.95,   18.50),
        ('SKU-1023', 'Barcode Scanner 2D Handheld',                26,   68.00,   72.00),
        ('SKU-1024', 'Wireless Presentation Clicker Remote',      109,   15.50,   14.00),
        ('SKU-1025', 'Fingerprint Reader USB Security Key',        83,   27.00,   23.50),
        ('SKU-1026', 'Document Shredder Micro-Cut 12-Sheet',       14,  145.00,  130.00),
        ('SKU-1027', 'Mini PC Intel N100 16GB RAM',                 9,  289.00,  265.00),
        ('SKU-1028', 'Power Bank 26800mAh 65W PD Fast Charge',    147,   38.00,   34.00),
        ('SKU-1029', 'Desk Organiser Bamboo with Drawers',        218,   16.00,   14.50),
        ('SKU-1030', 'Network Switch 8-Port Gigabit Unmanaged',    55,   22.50,   20.00),
        ('SKU-1031', 'Cat6 Ethernet Cable 25ft Flat',             390,    6.80,    6.20),
        ('SKU-1032', 'Smart Plug Wi-Fi 4-Pack 15A',               276,   12.00,   10.80),
        ('SKU-1033', 'Digital Microscope 1000x USB',               18,   75.00,   68.00),
        ('SKU-1034', 'USB Hub 7-Port 3.0 Self-Powered',           224,   14.50,   13.00),
        ('SKU-1035', 'Wrist Rest Ergonomic Memory Foam Pair',      341,    9.20,    8.50),
        ('SKU-1036', 'VoIP Conference Phone 360 Audio',            11,  198.00,  185.00),
        ('SKU-1037', 'Hard Drive Docking Station USB-C',           62,   38.50,   35.00),
        ('SKU-1038', 'Anti-Fatigue Mat Standing Desk 20x32in',    103,   28.00,   25.50),
        ('SKU-1039', 'Portable Projector Mini LED 200 Lumens',     24,  135.00,  122.00),
        ('SKU-1040', 'Whiteboard Markers Assorted 12-Pack',       512,    3.80,    3.50),
        ('SKU-1041', 'Thermal Receipt Printer 80mm',               17,   85.00,   92.00),
        ('SKU-1042', 'Network Attached Storage 2-Bay NAS',          7,  345.00,  320.00),
        ('SKU-1043', 'Soldering Station Digital 60W',              12,   62.00,   57.00),
        ('SKU-1044', 'PC Cleaning Kit Compressed Air + Brush',    288,    5.50,    5.00),
        ('SKU-1045', 'Wireless Number Pad Numpad Bluetooth',       96,   14.00,   12.50),
        ('SKU-1046', 'Colour Laser Toner Cyan High Yield',         51,   58.00,   52.00),
        ('SKU-1047', 'Smart Card Reader USB Compact',             134,   11.50,   10.20),
        ('SKU-1048', 'Monitor Calibration Device Colorimeter',      6,  175.00,  162.00),
        ('SKU-1049', 'HDMI Splitter 1-in-4-out 4K',               78,   18.00,   16.50),
        ('SKU-1050', 'Keyboard Wrist Rest Silicone Gel',          257,    7.80,    7.20),
        ('SKU-1051', 'Password Manager Hardware Token',            45,   32.00,   29.00),
        ('SKU-1052', 'USB Microphone Cardioid Condenser',          53,   48.00,   43.50),
        ('SKU-1053', 'Green Screen Fabric Chromakey 10x12ft',      29,   28.00,   25.00),
        ('SKU-1054', 'Cable Tester RJ45 Network Diagnostic',       37,   16.50,   15.00),
        ('SKU-1055', 'Electric Stapler Automatic 30-Sheet',        43,   22.00,   19.50),
        ('SKU-1056', 'Document Scanner Portable A4 Wi-Fi',         16,  118.00,  105.00),
        ('SKU-1057', 'LED Strip Lights 32ft Smart RGB',           198,   13.50,   12.00),
        ('SKU-1058', 'UPS Uninterruptible Power Supply 600VA',     21,   88.00,   82.00),
        ('SKU-1059', 'Touchscreen Display 10in POS System',         8,  225.00,  215.00),
    ]

    for r, (sku, desc, units, fifo_cost, repl_cost) in enumerate(inventory_items, 2):
        ws.cell(row=r, column=1, value=sku)
        ws.cell(row=r, column=2, value=desc)
        ws.cell(row=r, column=3, value=units)
        ws.cell(row=r, column=4, value=fifo_cost)
        ws.cell(row=r, column=5, value=repl_cost)
        # Columns F (6), G (7), H (8) intentionally left empty

    # Column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Inventory')
    print(f'  Rows: 1 header + 59 data rows (rows 2-60)')
    print(f'  Columns A-E filled; F, G, H intentionally empty')


create_initial()
