"""
Initial Setup: Create DeptMetrics sheet with 150 rows of department performance data.
Task ID: calc_pivot_090
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_090'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'DeptMetrics'

    # Headers
    headers = ['ID', 'Department', 'Metric', 'Value']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Target sums per department:
    # Engineering=45000, Product=38000, Design=28000, Marketing=22000, Sales=35000
    # Total = 168000
    # 150 rows, distribute as 30 per department
    dept_targets = {
        'Engineering': 45000,
        'Product': 38000,
        'Design': 28000,
        'Marketing': 22000,
        'Sales': 35000,
    }

    metrics = [
        'Revenue', 'Cost Savings', 'Productivity', 'Customer Satisfaction',
        'Quality Score', 'Efficiency', 'Output Volume', 'Error Rate Reduction',
        'Team Performance', 'Innovation Index'
    ]

    rows_per_dept = 30
    all_rows = []

    for dept, target_sum in dept_targets.items():
        # Generate 30 random values that sum to target_sum
        # Start with random proportions, then scale
        raw = [random.randint(100, 3000) for _ in range(rows_per_dept)]
        raw_sum = sum(raw)
        # Scale to target, rounding to integers
        scaled = [int(round(v / raw_sum * target_sum)) for v in raw]
        # Fix rounding error on last element
        diff = target_sum - sum(scaled)
        scaled[-1] += diff

        for i, val in enumerate(scaled):
            metric = metrics[i % len(metrics)]
            all_rows.append((dept, metric, val))

    # Shuffle to make it realistic (not grouped by department)
    random.shuffle(all_rows)

    for idx, (dept, metric, val) in enumerate(all_rows, 1):
        row = idx + 1  # row 2 onwards
        ws.cell(row=row, column=1, value=idx)          # ID
        ws.cell(row=row, column=2, value=dept)          # Department
        ws.cell(row=row, column=3, value=metric)        # Metric
        ws.cell(row=row, column=4, value=val)            # Value

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Verify sums
    dept_sums = {}
    for idx, (dept, metric, val) in enumerate(all_rows):
        dept_sums[dept] = dept_sums.get(dept, 0) + val
    print(f'Department sums: {dept_sums}')
    print(f'Grand total: {sum(dept_sums.values())}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
