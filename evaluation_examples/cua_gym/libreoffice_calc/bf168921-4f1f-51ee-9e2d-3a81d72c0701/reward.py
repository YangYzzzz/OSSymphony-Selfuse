"""
Reward Script: Create a pivot table with custom department sort order
Task ID: calc_pivot_090
Domain: libreoffice_calc
Scoring:
  Component 1: PivotTable sheet exists (0.15)
  Component 2: Correct headers (0.10)
  Component 3: All 5 departments with correct SUM values (0.35)
  Component 4: Custom sort order Engineering, Product, Design, Marketing, Sales (0.25)
  Component 5: Grand Total row with value 168000 (0.15)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_090'

# Expected custom order and values from task context
EXPECTED_ORDER = ['Engineering', 'Product', 'Design', 'Marketing', 'Sales']
EXPECTED_VALUES = {
    'Engineering': 45000,
    'Product': 38000,
    'Design': 28000,
    'Marketing': 22000,
    'Sales': 35000,
}
EXPECTED_GRAND_TOTAL = 168000


def persist_app_state(domain):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: PivotTable sheet exists (0.15 points)
    # This sheet does NOT exist in initial_env, only in golden_env.
    pivot_ws = None
    try:
        pivot_sheet_names = [s for s in wb.sheetnames if s != 'DeptMetrics']
        if len(pivot_sheet_names) >= 1:
            # Find a sheet that looks like a pivot table (has Department + value columns)
            for sn in pivot_sheet_names:
                ws_candidate = wb[sn]
                if ws_candidate.max_row >= 6 and ws_candidate.max_column >= 2:
                    pivot_ws = ws_candidate
                    break
            if pivot_ws is None:
                # Fallback: try any non-DeptMetrics sheet
                pivot_ws = wb[pivot_sheet_names[0]]
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    if pivot_ws is not None:
        print(f"PASS: Component 1 -- Pivot table sheet found: '{pivot_ws.title}' (0.15 pts)")
        total_score += 0.15
    else:
        print("FAIL: Component 1 -- No pivot table sheet found (only DeptMetrics exists)")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Component 2: Correct headers -- "Department" in A1-ish, value header in B1-ish (0.10 points)
    try:
        header_a = str(pivot_ws.cell(row=1, column=1).value or '').strip()
        header_b = str(pivot_ws.cell(row=1, column=2).value or '').strip()

        dept_header_ok = 'department' in header_a.lower()
        value_header_ok = 'value' in header_b.lower() or 'sum' in header_b.lower()

        if dept_header_ok and value_header_ok:
            print(f"PASS: Component 2 -- Headers correct: '{header_a}', '{header_b}' (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 -- Headers: A1='{header_a}', B1='{header_b}'. "
                  f"Expected 'Department' and 'Sum of Value' (or similar)")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: All 5 departments present with correct SUM values (0.35 points)
    # 0.07 per department
    try:
        found_depts = {}
        for row_idx in range(2, pivot_ws.max_row + 1):
            dept_val = pivot_ws.cell(row=row_idx, column=1).value
            sum_val = pivot_ws.cell(row=row_idx, column=2).value
            if dept_val and str(dept_val).strip() in EXPECTED_VALUES:
                found_depts[str(dept_val).strip()] = sum_val

        dept_score = 0.0
        for dept, expected_val in EXPECTED_VALUES.items():
            if dept in found_depts:
                actual_val = found_depts[dept]
                try:
                    if abs(float(actual_val) - expected_val) < 1.0:
                        print(f"  PASS: {dept} = {actual_val} (expected {expected_val})")
                        dept_score += 0.07
                    else:
                        print(f"  FAIL: {dept} = {actual_val} (expected {expected_val})")
                except (TypeError, ValueError):
                    print(f"  FAIL: {dept} has non-numeric value: {actual_val}")
            else:
                print(f"  FAIL: {dept} not found in pivot table")

        if dept_score > 0:
            print(f"PASS: Component 3 -- Department values ({dept_score:.2f} pts)")
            total_score += dept_score
        else:
            print("FAIL: Component 3 -- No departments with correct values found")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Custom sort order (0.25 points)
    # Departments must appear in order: Engineering, Product, Design, Marketing, Sales
    # This is the key task-introduced change -- NOT alphabetical order.
    try:
        actual_order = []
        for row_idx in range(2, pivot_ws.max_row + 1):
            dept_val = pivot_ws.cell(row=row_idx, column=1).value
            if dept_val and str(dept_val).strip() in EXPECTED_VALUES:
                actual_order.append(str(dept_val).strip())

        if actual_order == EXPECTED_ORDER:
            print(f"PASS: Component 4 -- Custom order correct: {actual_order} (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 4 -- Order mismatch. Got: {actual_order}, "
                  f"Expected: {EXPECTED_ORDER}")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # Component 5: Grand Total row with correct value 168000 (0.15 points)
    try:
        gt_label = None
        gt_value = None
        for row_idx in range(2, pivot_ws.max_row + 1):
            cell_a = pivot_ws.cell(row=row_idx, column=1).value
            cell_b = pivot_ws.cell(row=row_idx, column=2).value
            if cell_a and 'total' in str(cell_a).lower():
                gt_label = str(cell_a)
                gt_value = cell_b
                break

        if gt_label is not None and gt_value is not None:
            if abs(float(gt_value) - EXPECTED_GRAND_TOTAL) < 1.0:
                print(f"PASS: Component 5 -- Grand Total = {gt_value} (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 5 -- Grand Total = {gt_value}, "
                      f"expected {EXPECTED_GRAND_TOTAL}")
        else:
            print("FAIL: Component 5 -- No Grand Total row found")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
