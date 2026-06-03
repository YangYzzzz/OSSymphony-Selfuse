"""
FINAL REWARD SCRIPT - SUCCESS
Task: Transfer all numbers from 'Order ID' to 'Formatted Order ID' and add zeros at the beginning to make each one exactly 12 digits long. Complete this without touching irrelevant regions.
Generated: 2025-11-24 07:26:00
Status: success
Model: o3
Total Steps: 1
"""

import openpyxl
import os
import traceback
from openpyxl.utils import get_column_letter


def verify_task(file_path: str) -> float:
    """Verify that all numbers from 'Order ID' have been copied to
    'Formatted Order ID' as 12-digit zero-padded strings and that no
    unrelated data was changed.

    Returns a progressive score between 0.0 and 1.0 and prints detailed
    diagnostics for each verification step.
    """

    # Scoring weights (must sum to 1.0)
    WEIGHTS = {
        "formatted": 0.60,            # core requirement
        "order_id_unchanged": 0.10,   # ensure original data untouched
        "customer_unchanged": 0.15,   # ensure unrelated column intact
        "amount_unchanged": 0.15,     # ensure unrelated column intact
    }

    max_score = 1.0

    # Baseline (initial) data expected to remain unchanged in the workbook
    baseline_rows = [
        {"Order ID": "12345", "Customer": "Alpha Co.",   "Amount": "250"},
        {"Order ID": "987654321", "Customer": "Bravo LLC",  "Amount": "175.5"},
        {"Order ID": "55", "Customer": "Charlie",    "Amount": "98.15"},
        {"Order ID": "2021123001", "Customer": "Delta Inc",  "Amount": "600"},
        {"Order ID": "0", "Customer": "Echo GmbH",   "Amount": "12"},
    ]

    print(f"Starting verification for file: {file_path}\n")

    if not os.path.exists(file_path):
        print("✗ File not found. Task incomplete.")
        return 0.0

    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        if "Orders" not in wb.sheetnames:
            print("✗ Sheet 'Orders' not found.")
            return 0.0
        sheet = wb["Orders"]
        print("✓ Workbook and sheet loaded successfully")
    except Exception as e:
        print("✗ Error loading workbook:", e)
        traceback.print_exc()
        return 0.0

    # Map header names to column indices (1-based)
    headers = {}
    for col in range(1, sheet.max_column + 1):
        header_val = sheet.cell(row=1, column=col).value
        if header_val is not None:
            headers[str(header_val).strip()] = col

    required_cols = ["Order ID", "Formatted Order ID", "Customer", "Amount"]
    missing = [c for c in required_cols if c not in headers]
    if missing:
        print(f"✗ Missing expected columns: {missing}")
        return 0.0

    order_col     = headers["Order ID"]
    formatted_col = headers["Formatted Order ID"]
    customer_col  = headers["Customer"]
    amount_col    = headers["Amount"]

    total_rows         = len(baseline_rows)
    correct_formatted  = 0
    order_unchanged    = 0
    customer_unchanged = 0
    amount_unchanged   = 0

    print("\nVerifying each data row...\n")

    for idx, baseline in enumerate(baseline_rows, start=2):  # data starts on row 2
        order_val     = sheet.cell(row=idx, column=order_col).value
        formatted_val = sheet.cell(row=idx, column=formatted_col).value
        customer_val  = sheet.cell(row=idx, column=customer_col).value
        amount_val    = sheet.cell(row=idx, column=amount_col).value

        # Standardise string representations for reliable comparison
        order_str    = "" if order_val is None else str(order_val).strip()
        formatted_str = "" if formatted_val is None else str(formatted_val).lstrip("'").strip()
        customer_str = "" if customer_val is None else str(customer_val).strip()
        amount_str   = "" if amount_val is None else str(amount_val).strip()

        baseline_order    = baseline["Order ID"]
        baseline_customer = baseline["Customer"]
        baseline_amount   = baseline["Amount"]

        # 1. Ensure original Order ID is untouched
        if order_str == baseline_order:
            order_unchanged += 1
        else:
            print(f"Row {idx}: ✗ Order ID changed (expected '{baseline_order}', found '{order_str}')")

        # 2. Ensure Customer unchanged
        if customer_str == baseline_customer:
            customer_unchanged += 1
        else:
            print(f"Row {idx}: ✗ Customer modified (expected '{baseline_customer}', found '{customer_str}')")

        # 3. Ensure Amount unchanged (allow numeric equivalence)
        amount_ok = False
        try:
            amount_ok = float(amount_str) == float(baseline_amount)
        except Exception:
            pass
        if amount_str == baseline_amount or amount_ok:
            amount_unchanged += 1
        else:
            print(f"Row {idx}: ✗ Amount modified (expected '{baseline_amount}', found '{amount_str}')")

        # 4. Verify correct 12-digit zero-padded formatted ID
        expected_formatted = baseline_order.zfill(12)

        # Accept direct text match
        if formatted_str == expected_formatted:
            correct_formatted += 1
            continue

        # Accept a VALID formula that would evaluate to the same result
        if isinstance(formatted_val, str) and formatted_val.startswith("="):
            col_letter = get_column_letter(order_col)
            cell_ref   = f"{col_letter}{idx}"
            formula_up = formatted_val.upper()
            # basic heuristic: formula references the order ID cell and uses TEXT/RIGHT/etc.
            if cell_ref.upper() in formula_up and any(func in formula_up for func in ("TEXT(", "RIGHT(", "CONCAT", "REPT(")):
                print(f"Row {idx}: ✓ Valid formula detected for formatted ID")
                correct_formatted += 1
            else:
                print(f"Row {idx}: ✗ Formatted ID incorrect (expected '{expected_formatted}', found formula '{formatted_val}')")
        else:
            print(f"Row {idx}: ✗ Formatted ID incorrect (expected '{expected_formatted}', found '{formatted_str}')")

    # --- Scoring ----------------------------------------------------------------
    formatted_score = (correct_formatted  / total_rows) * WEIGHTS["formatted"]
    order_score     = (order_unchanged    / total_rows) * WEIGHTS["order_id_unchanged"]
    customer_score  = (customer_unchanged / total_rows) * WEIGHTS["customer_unchanged"]
    amount_score    = (amount_unchanged   / total_rows) * WEIGHTS["amount_unchanged"]

    total_score = round(min(formatted_score + order_score + customer_score + amount_score, max_score), 4)

    # --- Reporting --------------------------------------------------------------
    print("\nScore details:")
    print(f" - Formatted IDs correct : {correct_formatted}/{total_rows} -> {formatted_score:0.2f} points")
    print(f" - Order IDs unchanged   : {order_unchanged}/{total_rows} -> {order_score:0.2f} points")
    print(f" - Customers unchanged   : {customer_unchanged}/{total_rows} -> {customer_score:0.2f} points")
    print(f" - Amounts unchanged     : {amount_unchanged}/{total_rows} -> {amount_score:0.2f} points")
    print(f"Total score: {total_score:0.2f} / {max_score}\n")

    return total_score


# ------------------------------------------------------------------------------
# When executed directly, run verification and print final reward
# ------------------------------------------------------------------------------
if __name__ == "__main__":
    path = "/home/user/transfer_all_numbers_from_order_id_to_formatted_order_id_and_add_zeros_at_the_beginning_to_make_each.xlsx"
    reward = verify_task(path)
    print(f"REWARD: {reward}")
