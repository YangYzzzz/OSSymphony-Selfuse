"""
Initial Setup: Project cost estimation worksheet with raw data
Task ID: calc_gpm_052
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_052'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'CostEst'

    # Row 1: Plain title (no merge, no formatting - agent must format)
    ws['A1'] = 'Project Cost Estimation - Website Redesign'

    # Row 3: Plain headers (no formatting - agent must format)
    headers = ['Phase', 'Task', 'Hours', 'Rate', 'Cost', 'Contingency']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)

    # Data rows 4-17: raw data only
    # Cost (E) and Contingency (F) left EMPTY - agent must add formulas
    # Using exactly the tasks from context, padded to fill rows 4-17
    data = [
        # Discovery phase (rows 4-6)
        ['Discovery', 'Stakeholder Interviews', 16, 150],
        ['Discovery', 'User Research', 24, 125],
        ['Discovery', 'Requirements Doc', 12, 150],
        # Design phase (rows 7-9)
        ['Design', 'Wireframes', 20, 130],
        ['Design', 'Visual Design', 32, 130],
        ['Design', 'Prototyping', 16, 130],
        # Development phase (rows 10-13)
        ['Development', 'Frontend', 80, 140],
        ['Development', 'Backend', 60, 140],
        ['Development', 'CMS Setup', 24, 120],
        ['Development', 'API Integration', 32, 140],
        # Testing phase (rows 14-17)
        ['Testing', 'QA Testing', 24, 110],
        ['Testing', 'UAT Support', 16, 110],
        ['Testing', 'Performance Testing', 20, 120],
        ['Testing', 'Bug Fix Support', 16, 130],
    ]

    for r, row_data in enumerate(data, 4):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Row 18-19: Labels only (no formulas, no formatting)
    ws['A18'] = 'Subtotal'
    ws['A19'] = 'Total with Contingency'

    # Set reasonable column widths for readability
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
