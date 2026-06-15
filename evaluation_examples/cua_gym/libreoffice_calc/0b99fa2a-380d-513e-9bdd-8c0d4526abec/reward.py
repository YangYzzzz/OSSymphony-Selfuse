"""
Reward Script: Build a project cost estimation worksheet with itemized costs,
contingency calculation, and formatted summary.
Task ID: calc_gpm_052
Domain: libreoffice_calc
Scoring:
  C1: Cost & Contingency formulas in E/F columns (0.20)
  C2: Merged cells for title and phases (0.15)
  C3: Title formatting - bold 14pt centered orange fill white text (0.10)
  C4: Header row formatting - bold orange fill white text (0.10)
  C5: Summary row formulas (SUM, Total) with E19 double underline (0.15)
  C6: Phase Summary section with SUMIF formulas (0.10)
  C7: Pie chart present with title (0.10)
  C8: Number format $#,##0 and conditional formatting (0.10)
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_052'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists
    if 'CostEst' not in wb.sheetnames:
        print("FAIL: Sheet 'CostEst' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['CostEst']

    # Component 1: Cost & Contingency formulas in E and F columns (0.20 points)
    # Initial file has NO formulas in E or F — only raw data in A-D
    try:
        formula_count = 0
        expected_formulas = 0
        for row in range(4, 18):  # rows 4-17 data rows
            e_val = ws.cell(row=row, column=5).value  # column E
            f_val = ws.cell(row=row, column=6).value  # column F
            if e_val is not None:
                expected_formulas += 1
                if isinstance(e_val, str) and '=' in e_val and 'C' in e_val.upper() and 'D' in e_val.upper():
                    formula_count += 1
            if f_val is not None:
                expected_formulas += 1
                if isinstance(f_val, str) and '=' in f_val and '0.15' in f_val:
                    formula_count += 1

        if expected_formulas > 0 and formula_count >= 20:
            print(f"PASS: Component 1 - Cost/Contingency formulas found ({formula_count} formulas) (0.20 pts)")
            total_score += 0.20
        elif formula_count >= 10:
            print(f"PARTIAL: Component 1 - Some formulas found ({formula_count}) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 - Expected E/F formulas, found {formula_count} of {expected_formulas}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Merged cells for title and phase cells (0.15 points)
    # Initial file has NO merged cells at all
    try:
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        has_title_merge = any('A1' in r and 'F1' in r for r in merged_ranges)
        # Check phase merges - at least some phase cells should be merged in column A
        phase_merges = sum(1 for r in merged_ranges if r.startswith('A') and r[1:].split(':')[0].isdigit())
        # H3:I3 merge
        has_summary_header_merge = any('H3' in r and 'I3' in r for r in merged_ranges)

        merge_score = 0.0
        if has_title_merge:
            merge_score += 0.05
        if phase_merges >= 3:
            merge_score += 0.05
        if has_summary_header_merge:
            merge_score += 0.05

        if merge_score > 0:
            print(f"PASS: Component 2 - Merged cells: title={has_title_merge}, phase_merges={phase_merges}, summary_header={has_summary_header_merge} ({merge_score} pts)")
            total_score += merge_score
        else:
            print(f"FAIL: Component 2 - No expected merged cells found. Ranges: {merged_ranges}")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Title formatting - A1 bold 14pt centered, orange fill, white font (0.10 points)
    # Initial file A1 has no formatting at all
    try:
        a1 = ws['A1']
        if isinstance(a1, MergedCell):
            # If A1 is merged, it's not the top-left; skip
            print("FAIL: Component 3 - A1 is a MergedCell (unexpected)")
        else:
            title_score = 0.0
            is_bold = a1.font.bold is True
            is_14pt = a1.font.size is not None and abs(a1.font.size - 14.0) < 1
            is_centered = a1.alignment.horizontal == 'center'

            # Check fill color - expect dark orange (CC6600)
            fill_rgb = None
            try:
                fill_rgb = a1.fill.fgColor.rgb
            except:
                pass
            has_orange_fill = fill_rgb is not None and 'CC6600' in str(fill_rgb).upper()

            # Check white font
            font_rgb = None
            try:
                font_rgb = a1.font.color.rgb
            except:
                pass
            has_white_font = font_rgb is not None and 'FFFFFF' in str(font_rgb).upper()

            checks_passed = sum([is_bold, is_14pt, is_centered, has_orange_fill, has_white_font])
            if checks_passed >= 4:
                title_score = 0.10
            elif checks_passed >= 2:
                title_score = 0.05

            if title_score > 0:
                print(f"PASS: Component 3 - Title formatting ({checks_passed}/5 checks) ({title_score} pts)")
                total_score += title_score
            else:
                print(f"FAIL: Component 3 - Title formatting: bold={is_bold}, 14pt={is_14pt}, center={is_centered}, orange_fill={has_orange_fill}(rgb={fill_rgb}), white_font={has_white_font}(rgb={font_rgb})")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Header row formatting - A3:F3 bold, orange fill, white text (0.10 points)
    # Initial file has no header formatting
    try:
        header_bold_count = 0
        header_fill_count = 0
        for col in range(1, 7):  # A3 to F3
            cell = ws.cell(row=3, column=col)
            if cell.font.bold is True:
                header_bold_count += 1
            try:
                fill_rgb = cell.fill.fgColor.rgb
                if fill_rgb and 'CC6600' in str(fill_rgb).upper():
                    header_fill_count += 1
            except:
                pass

        if header_bold_count >= 5 and header_fill_count >= 5:
            print(f"PASS: Component 4 - Headers formatted: {header_bold_count}/6 bold, {header_fill_count}/6 orange fill (0.10 pts)")
            total_score += 0.10
        elif header_bold_count >= 3 or header_fill_count >= 3:
            print(f"PARTIAL: Component 4 - Partial header formatting: {header_bold_count}/6 bold, {header_fill_count}/6 fill (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4 - Header formatting: {header_bold_count}/6 bold, {header_fill_count}/6 orange fill")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Summary formulas - E18=SUM, F18=SUM, E19=E18+F18, E19 double underline (0.15 points)
    # Initial file has NO formulas in E18/F18/E19 and no styling
    try:
        summary_score = 0.0

        e18 = ws['E18'].value
        f18 = ws['F18'].value
        e19 = ws['E19'].value

        has_e18_sum = isinstance(e18, str) and 'SUM' in e18.upper()
        has_f18_sum = isinstance(f18, str) and 'SUM' in f18.upper()
        has_e19_total = isinstance(e19, str) and '=' in e19

        if has_e18_sum and has_f18_sum:
            summary_score += 0.05
        if has_e19_total:
            summary_score += 0.05

        # Check E19 double underline and bold
        e19_cell = ws['E19']
        e19_underline = e19_cell.font.underline
        e19_bold = e19_cell.font.bold is True
        if e19_underline == 'double' and e19_bold:
            summary_score += 0.05

        if summary_score > 0:
            print(f"PASS: Component 5 - Summary: E18_SUM={has_e18_sum}, F18_SUM={has_f18_sum}, E19_total={has_e19_total}, E19_underline={e19_underline}, E19_bold={e19_bold} ({summary_score} pts)")
            total_score += summary_score
        else:
            print(f"FAIL: Component 5 - E18={e18}, F18={f18}, E19={e19}, underline={e19_underline}")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    # Component 6: Phase Summary with SUMIF formulas (0.10 points)
    # Initial has NO Phase Summary section at all (no H3 label, no SUMIF formulas)
    try:
        h3_val = ws['H3'].value
        if isinstance(ws['H3'], MergedCell):
            # Try to find the value from merged range top-left
            h3_val = None
            for mr in ws.merged_cells.ranges:
                if 'H3' in str(mr):
                    tl = str(mr).split(':')[0]
                    h3_val = ws[tl].value
                    break

        has_phase_summary_header = h3_val is not None and 'Phase Summary' in str(h3_val)

        sumif_count = 0
        for row in range(4, 8):
            i_val = ws.cell(row=row, column=9).value  # column I
            if isinstance(i_val, str) and 'SUMIF' in i_val.upper():
                sumif_count += 1

        phase_summary_score = 0.0
        if has_phase_summary_header and sumif_count >= 3:
            phase_summary_score = 0.10
        elif has_phase_summary_header or sumif_count >= 2:
            phase_summary_score = 0.05

        if phase_summary_score > 0:
            print(f"PASS: Component 6 - Phase Summary: header={has_phase_summary_header}, SUMIF_count={sumif_count} ({phase_summary_score} pts)")
            total_score += phase_summary_score
        else:
            print(f"FAIL: Component 6 - Phase Summary: header={has_phase_summary_header}, SUMIF_count={sumif_count}")
    except Exception as e:
        print(f"ERROR: Component 6 - {e}")

    # Component 7: Pie chart present (0.10 points)
    # Initial file has NO charts
    try:
        charts = ws._charts
        pie_charts = [c for c in charts if c.__class__.__name__ == 'PieChart']
        chart_title = None
        if len(pie_charts) > 0:
            try:
                for p in pie_charts[0].title.tx.rich.paragraphs:
                    for r in p.r:
                        chart_title = r.t
            except:
                chart_title = str(pie_charts[0].title) if pie_charts[0].title else None

        if len(pie_charts) > 0:
            print(f"PASS: Component 7 - Pie chart found, title='{chart_title}' (0.10 pts)")
            total_score += 0.10
        elif len(charts) > 0:
            print(f"PARTIAL: Component 7 - Chart found but not PieChart, types={[c.__class__.__name__ for c in charts]} (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 7 - No charts found")
    except Exception as e:
        print(f"ERROR: Component 7 - {e}")

    # Component 8: Number format $#,##0 on cost columns AND conditional formatting (0.10 points)
    # Initial file has General format and no conditional formatting
    try:
        fmt_score = 0.0

        # Check number format on D, E, F columns (data rows)
        dollar_fmt_count = 0
        for row in range(4, 18):
            for col in [4, 5, 6]:  # D, E, F
                nf = ws.cell(row=row, column=col).number_format
                if nf and '$' in str(nf):
                    dollar_fmt_count += 1

        if dollar_fmt_count >= 30:
            fmt_score += 0.05

        # Check conditional formatting exists
        cf_count = 0
        data_bar_count = 0
        cell_is_count = 0
        for cf in ws.conditional_formatting:
            for rule in cf.rules:
                cf_count += 1
                if rule.type == 'dataBar':
                    data_bar_count += 1
                if rule.type == 'cellIs':
                    cell_is_count += 1

        if data_bar_count > 0 or cell_is_count > 0:
            fmt_score += 0.05

        if fmt_score > 0:
            print(f"PASS: Component 8 - Number format: {dollar_fmt_count} cells with $, CF: dataBar={data_bar_count}, cellIs={cell_is_count} ({fmt_score} pts)")
            total_score += fmt_score
        else:
            print(f"FAIL: Component 8 - Number format: {dollar_fmt_count} cells with $, CF count={cf_count}")
    except Exception as e:
        print(f"ERROR: Component 8 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
