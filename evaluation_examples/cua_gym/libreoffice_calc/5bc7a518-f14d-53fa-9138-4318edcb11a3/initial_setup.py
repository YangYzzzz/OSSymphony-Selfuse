"""
Initial Setup: Multi-tier pricing calculator with quantity breaks, customer tier discounts, and contract length discounts
Task ID: calc_sales_060
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_060'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Header styling ---
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # === Sheet 1: PriceCalc ===
    ws1 = wb.active
    ws1.title = 'PriceCalc'

    headers = ['Order', 'Base Price', 'Qty', 'Customer Tier', 'Contract (months)',
               'Qty Discount', 'Tier Discount', 'Contract Discount', 'Final Unit Price', 'Total']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows (columns A-E only; F-J left empty for agent to fill with formulas)
    data = [
        ['O1', 100, 250, 'Gold', 36],
        ['O2', 100, 50, 'Silver', 12],
        ['O3', 100, 500, 'Platinum', 24],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 2:
                cell.number_format = '$#,##0.00'

    # Also add borders to the empty F-J cells so it looks like a real form
    for r in range(2, 5):
        for c in range(6, 11):
            cell = ws1.cell(row=r, column=c)
            cell.border = thin_border
            if c in (6, 7, 8):
                cell.number_format = '0.00%'
            elif c == 9:
                cell.number_format = '$#,##0.00'
            elif c == 10:
                cell.number_format = '$#,##0.00'

    # Set column widths for readability
    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 14
    ws1.column_dimensions['C'].width = 8
    ws1.column_dimensions['D'].width = 16
    ws1.column_dimensions['E'].width = 18
    ws1.column_dimensions['F'].width = 14
    ws1.column_dimensions['G'].width = 14
    ws1.column_dimensions['H'].width = 18
    ws1.column_dimensions['I'].width = 16
    ws1.column_dimensions['J'].width = 14

    # === Sheet 2: QtyBreaks ===
    ws2 = wb.create_sheet('QtyBreaks')
    qty_headers = ['Min', 'Discount']
    for col, h in enumerate(qty_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    qty_data = [
        [1, 0],
        [100, 0.05],
        [250, 0.10],
        [500, 0.15],
    ]
    for r, row_data in enumerate(qty_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 2:
                cell.number_format = '0%'

    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 12

    # === Sheet 3: TierDisc ===
    ws3 = wb.create_sheet('TierDisc')
    tier_headers = ['Tier', 'Discount']
    for col, h in enumerate(tier_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    tier_data = [
        ['Silver', 0.05],
        ['Gold', 0.10],
        ['Platinum', 0.15],
    ]
    for r, row_data in enumerate(tier_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 2:
                cell.number_format = '0%'

    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 12

    # === Sheet 4: ContractDisc ===
    ws4 = wb.create_sheet('ContractDisc')
    contract_headers = ['Months', 'Discount']
    for col, h in enumerate(contract_headers, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    contract_data = [
        [12, 0],
        [24, 0.05],
        [36, 0.10],
    ]
    for r, row_data in enumerate(contract_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws4.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 2:
                cell.number_format = '0%'

    ws4.column_dimensions['A'].width = 12
    ws4.column_dimensions['B'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
