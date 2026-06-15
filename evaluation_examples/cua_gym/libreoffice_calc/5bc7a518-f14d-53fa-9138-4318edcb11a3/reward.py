"""
Reward Script: Multi-tier pricing calculator with VLOOKUP formulas
Task ID: calc_sales_060
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): VLOOKUP formulas in F2:F4 (Qty Discount lookups)
  Component 2 (0.20): VLOOKUP formulas in G2:G4 (Tier Discount lookups)
  Component 3 (0.20): VLOOKUP formulas in H2:H4 (Contract Discount lookups)
  Component 4 (0.15): Compound formulas in I2:I4 (Final Unit Price)
  Component 5 (0.15): Formulas in J2:J4 (Total = I * C)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_060'


def normalize_formula(f):
    """Normalize a formula for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def is_vlookup_referencing(formula, lookup_col, sheet_range):
    """
    Check if formula is a VLOOKUP that references the expected lookup column
    and sheet range. Flexible about exact syntax variations.
    """
    f = normalize_formula(formula)
    if not f.startswith('=VLOOKUP('):
        return False
    # Check it references the expected sheet range (e.g., QTYBREAKS!A:B)
    if sheet_range.upper().replace(' ', '') in f:
        return True
    # Also accept range variants like QtyBreaks!A1:B5, QtyBreaks!$A:$B etc.
    sheet_name = sheet_range.split('!')[0].upper()
    if sheet_name in f and 'VLOOKUP' in f:
        return True
    return False


def has_multiplication_pattern(formula, components):
    """
    Check if formula multiplies the expected cell references.
    E.g., =B2*(1-F2)*(1-G2)*(1-H2)
    """
    f = normalize_formula(formula)
    if not f.startswith('='):
        return False
    # Check that all expected components appear in the formula
    for comp in components:
        if comp.upper() not in f:
            return False
    return True


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: PriceCalc sheet must exist
    if 'PriceCalc' not in wb.sheetnames:
        print("CRITICAL: 'PriceCalc' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['PriceCalc']

    # Component 1: VLOOKUP formulas in F2:F4 for Qty Discount (0.30 points)
    # These should be VLOOKUPs referencing the QtyBreaks sheet
    try:
        comp1_score = 0.0
        rows_to_check = [2, 3, 4]
        per_row = 0.10
        for r in rows_to_check:
            cell_val = ws.cell(row=r, column=6).value  # Column F
            if cell_val is not None and is_vlookup_referencing(str(cell_val), 'C', 'QtyBreaks!A:B'):
                print(f"PASS: F{r} has VLOOKUP to QtyBreaks: {cell_val}")
                comp1_score += per_row
            else:
                print(f"FAIL: F{r} expected VLOOKUP to QtyBreaks, found: {cell_val}")
        if comp1_score > 0:
            total_score += comp1_score
            print(f"Component 1 subtotal: {comp1_score:.2f}/0.30")
    except Exception as e:
        print(f"ERROR: Component 1 (Qty Discount VLOOKUPs) - {e}")

    # Component 2: VLOOKUP formulas in G2:G4 for Tier Discount (0.20 points)
    try:
        comp2_score = 0.0
        per_row = round(0.20 / 3, 4)
        for r in rows_to_check:
            cell_val = ws.cell(row=r, column=7).value  # Column G
            if cell_val is not None and is_vlookup_referencing(str(cell_val), 'D', 'TierDisc!A:B'):
                print(f"PASS: G{r} has VLOOKUP to TierDisc: {cell_val}")
                comp2_score += per_row
            else:
                print(f"FAIL: G{r} expected VLOOKUP to TierDisc, found: {cell_val}")
        # Cap at 0.20
        comp2_score = min(comp2_score, 0.20)
        if comp2_score > 0:
            total_score += comp2_score
            print(f"Component 2 subtotal: {comp2_score:.2f}/0.20")
    except Exception as e:
        print(f"ERROR: Component 2 (Tier Discount VLOOKUPs) - {e}")

    # Component 3: VLOOKUP formulas in H2:H4 for Contract Discount (0.20 points)
    try:
        comp3_score = 0.0
        per_row = round(0.20 / 3, 4)
        for r in rows_to_check:
            cell_val = ws.cell(row=r, column=8).value  # Column H
            if cell_val is not None and is_vlookup_referencing(str(cell_val), 'E', 'ContractDisc!A:B'):
                print(f"PASS: H{r} has VLOOKUP to ContractDisc: {cell_val}")
                comp3_score += per_row
            else:
                print(f"FAIL: H{r} expected VLOOKUP to ContractDisc, found: {cell_val}")
        comp3_score = min(comp3_score, 0.20)
        if comp3_score > 0:
            total_score += comp3_score
            print(f"Component 3 subtotal: {comp3_score:.2f}/0.20")
    except Exception as e:
        print(f"ERROR: Component 3 (Contract Discount VLOOKUPs) - {e}")

    # Component 4: Compound formulas in I2:I4 (0.15 points)
    # Expected pattern: =B{r}*(1-F{r})*(1-G{r})*(1-H{r})
    try:
        comp4_score = 0.0
        per_row = 0.05
        for r in rows_to_check:
            cell_val = ws.cell(row=r, column=9).value  # Column I
            if cell_val is not None:
                f = normalize_formula(str(cell_val))
                # Check that it references B, F, G, H columns for this row
                expected_refs = [f'B{r}', f'F{r}', f'G{r}', f'H{r}']
                if has_multiplication_pattern(str(cell_val), expected_refs):
                    print(f"PASS: I{r} has compound formula: {cell_val}")
                    comp4_score += per_row
                else:
                    print(f"FAIL: I{r} expected compound formula with {expected_refs}, found: {cell_val}")
            else:
                print(f"FAIL: I{r} is empty, expected compound formula")
        if comp4_score > 0:
            total_score += comp4_score
            print(f"Component 4 subtotal: {comp4_score:.2f}/0.15")
    except Exception as e:
        print(f"ERROR: Component 4 (Final Unit Price formulas) - {e}")

    # Component 5: Formulas in J2:J4 (0.15 points)
    # Expected pattern: =I{r}*C{r}
    try:
        comp5_score = 0.0
        per_row = 0.05
        for r in rows_to_check:
            cell_val = ws.cell(row=r, column=10).value  # Column J
            if cell_val is not None:
                f = normalize_formula(str(cell_val))
                # Check it references I and C for this row
                if f.startswith('=') and f'I{r}' in f.upper() and f'C{r}' in f.upper():
                    print(f"PASS: J{r} has total formula: {cell_val}")
                    comp5_score += per_row
                else:
                    print(f"FAIL: J{r} expected formula with I{r}*C{r}, found: {cell_val}")
            else:
                print(f"FAIL: J{r} is empty, expected total formula")
        if comp5_score > 0:
            total_score += comp5_score
            print(f"Component 5 subtotal: {comp5_score:.2f}/0.15")
    except Exception as e:
        print(f"ERROR: Component 5 (Total formulas) - {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
