"""
Initial Setup: Merge and center cells A1:F1 to create a title cell
Task ID: calc_gfl_021
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_021'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: January ---
    ws = wb.active
    ws.title = 'January'

    # Row 1: Title (NOT merged, NOT centered - that's the task)
    ws.cell(row=1, column=1, value='Monthly Sales Performance Report - January 2024')
    # B1:F1 are empty (title overflows visually)

    # Row 2: Headers
    headers = ['Salesperson', 'Region', 'Target', 'Actual', 'Variance', 'Achievement%']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True)

    # Rows 3-25: 23 records of monthly sales data
    sales_data = [
        ['Sarah Chen', 'Northeast', 45000, 48230, 3230, 107.2],
        ['Marcus Johnson', 'Southeast', 52000, 49870, -2130, 95.9],
        ['Emily Rodriguez', 'West', 38000, 41560, 3560, 109.4],
        ['David Kim', 'Midwest', 41000, 38750, -2250, 94.5],
        ['Jennifer Walsh', 'Northeast', 47000, 52100, 5100, 110.9],
        ['Robert Taylor', 'Southeast', 55000, 54200, -800, 98.5],
        ['Lisa Nakamura', 'West', 43000, 45890, 2890, 106.7],
        ['Michael Brown', 'Midwest', 39000, 36420, -2580, 93.4],
        ['Amanda Foster', 'Northeast', 50000, 53670, 3670, 107.3],
        ['James Wilson', 'Southeast', 48000, 47100, -900, 98.1],
        ['Rachel Green', 'West', 42000, 44350, 2350, 105.6],
        ['Thomas Martinez', 'Midwest', 37000, 39800, 2800, 107.6],
        ['Nicole Adams', 'Northeast', 46000, 48900, 2900, 106.3],
        ['Christopher Lee', 'Southeast', 51000, 49500, -1500, 97.1],
        ['Stephanie Clark', 'West', 40000, 42780, 2780, 107.0],
        ['Daniel Harris', 'Midwest', 44000, 41200, -2800, 93.6],
        ['Michelle Wang', 'Northeast', 49000, 51340, 2340, 104.8],
        ['Andrew Scott', 'Southeast', 53000, 55100, 2100, 103.9],
        ['Laura Bennett', 'West', 36000, 38450, 2450, 106.8],
        ['Kevin O\'Brien', 'Midwest', 42000, 40100, -1900, 95.5],
        ['Patricia Hughes', 'Northeast', 47500, 49200, 1700, 103.6],
        ['Steven Garcia', 'Southeast', 50500, 48700, -1800, 96.4],
        ['Karen Phillips', 'West', 41500, 43900, 2400, 105.8],
    ]

    for r, row_data in enumerate(sales_data, 3):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 16

    # Format number columns
    for r in range(3, 26):
        ws.cell(row=r, column=3).number_format = '#,##0'
        ws.cell(row=r, column=4).number_format = '#,##0'
        ws.cell(row=r, column=5).number_format = '#,##0'
        ws.cell(row=r, column=6).number_format = '0.0'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
