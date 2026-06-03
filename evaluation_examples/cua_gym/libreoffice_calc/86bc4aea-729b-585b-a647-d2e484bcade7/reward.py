"""
Reward Script: Merge and center cells A1:F1 to create a single title cell
Task ID: calc_gfl_021
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Cells A1:F1 are merged into a single range
  Component 2 (0.3): A1 has horizontal alignment "center"
  Component 3 (0.3): Title text preserved in merged cell A1
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_021'

EXPECTED_TITLE = 'Monthly Sales Performance Report - January 2024'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Access the January sheet
    try:
        ws = wb['January']
    except KeyError:
        print("CRITICAL: Sheet 'January' not found")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Cells A1:F1 are merged into a single range (0.4 points)
    # This FAILS on initial (no merged ranges) and PASSES on golden (A1:F1 merged)
    try:
        merged_ranges = list(ws.merged_cells.ranges)
        # Check if A1:F1 is among the merged ranges
        a1_f1_merged = any(
            mr.min_row == 1 and mr.max_row == 1 and mr.min_col == 1 and mr.max_col == 6
            for mr in merged_ranges
        )
        if a1_f1_merged:
            print(f"PASS: Component 1 — A1:F1 is merged (0.4 pts)")
            total_score += 0.4
        else:
            # Also check if B1-F1 are MergedCell objects as a secondary indicator
            b1_merged = isinstance(ws['B1'], MergedCell)
            f1_merged = isinstance(ws['F1'], MergedCell)
            if b1_merged and f1_merged:
                # Merged but maybe range coords differ slightly
                print(f"PASS: Component 1 — B1 and F1 are MergedCell objects, merge detected (0.4 pts)")
                total_score += 0.4
            else:
                print(f"FAIL: Component 1 — A1:F1 not merged. Merged ranges: {merged_ranges}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: A1 horizontal alignment is "center" (0.3 points)
    # This FAILS on initial (alignment is None/general) and PASSES on golden (center)
    try:
        alignment = ws['A1'].alignment
        h_align = alignment.horizontal
        if h_align == 'center':
            print(f"PASS: Component 2 — A1 horizontal alignment is 'center' (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — A1 horizontal alignment is '{h_align}', expected 'center'")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Title text preserved in A1 after merge (0.3 points)
    # This checks that the merge operation preserved the original title text.
    # On initial_env, A1 has the title BUT cells are not merged, so we gate this
    # behind Component 1 passing (merge must exist) to avoid awarding points pre-task.
    try:
        a1_value = ws['A1'].value
        # Gate: only score if cells are actually merged (task was performed)
        has_merge = any(
            mr.min_row == 1 and mr.max_row == 1 and mr.min_col == 1 and mr.max_col >= 6
            for mr in ws.merged_cells.ranges
        ) or isinstance(ws['B1'], MergedCell)

        if has_merge and a1_value and str(a1_value).strip() == EXPECTED_TITLE:
            print(f"PASS: Component 3 — Title text preserved: '{a1_value}' (0.3 pts)")
            total_score += 0.3
        elif not has_merge:
            print(f"FAIL: Component 3 — Cells not merged, so title preservation not scored")
        else:
            print(f"FAIL: Component 3 — A1 value is '{a1_value}', expected '{EXPECTED_TITLE}'")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
