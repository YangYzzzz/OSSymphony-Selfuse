"""
Initial Setup: Tax Estimation Worksheet
Task ID: calc_wf_047
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_047'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tax Estimate"

    # Column widths
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18

    # --- Title ---
    ws["A1"] = "2024 Federal Tax Estimate"
    ws["A1"].font = Font(size=14, bold=True)

    # =====================
    # INCOME SECTION (rows 3-10)
    # =====================
    ws["A3"] = "INCOME"
    ws["A3"].font = Font(bold=True, underline="single")

    ws["A4"] = "W-2 Wages (Primary Job)"
    ws["B4"] = 87500.00

    ws["A5"] = "W-2 Wages (Spouse)"
    ws["B5"] = 62400.00

    ws["A6"] = "Freelance / Self-Employment Income"
    ws["B6"] = 14200.00

    ws["A7"] = "Interest Income (Savings & CDs)"
    ws["B7"] = 1830.50

    ws["A8"] = "Dividend Income (Brokerage)"
    ws["B8"] = 3475.00

    ws["A9"] = "Capital Gains (Stock Sales)"
    ws["B9"] = 5620.00

    ws["A10"] = "Other Income (Rental)"
    ws["B10"] = 8400.00

    # =====================
    # ADJUSTMENTS SECTION (rows 12-16)
    # =====================
    ws["A12"] = "ADJUSTMENTS TO INCOME"
    ws["A12"].font = Font(bold=True, underline="single")

    ws["A13"] = "IRA Contribution"
    ws["B13"] = 6500.00

    ws["A14"] = "Student Loan Interest"
    ws["B14"] = 2500.00

    ws["A15"] = "Self-Employment Tax (50%)"
    ws["B15"] = 1003.10

    ws["A16"] = "HSA Contribution"
    ws["B16"] = 3850.00

    # =====================
    # AGI (row 18) - NO FORMULA in initial
    # =====================
    ws["A18"] = "ADJUSTED GROSS INCOME (AGI)"
    ws["A18"].font = Font(bold=True)
    # B18 left empty - agent must add formula

    # =====================
    # DEDUCTIONS SECTION (rows 20-28)
    # =====================
    ws["A20"] = "DEDUCTIONS"
    ws["A20"].font = Font(bold=True, underline="single")

    ws["A21"] = "Standard Deduction (MFJ 2024)"
    ws["B21"] = 29200.00

    ws["A22"] = "--- Itemized Deductions ---"
    ws["A22"].font = Font(italic=True)

    ws["A23"] = "Mortgage Interest"
    ws["B23"] = 14280.00

    ws["A24"] = "Charitable Contributions"
    ws["B24"] = 6750.00

    ws["A25"] = "State & Local Taxes (SALT, capped)"
    ws["B25"] = 10000.00

    ws["A26"] = "Medical Expenses (above 7.5% AGI)"
    ws["B26"] = 0.00

    ws["A27"] = "Total Itemized Deductions"
    # B27 left empty - agent must add formula

    ws["A28"] = "Deduction Used (Greater of Std/Itemized)"
    ws["A28"].font = Font(bold=True)
    # B28 left empty - agent must add formula

    # =====================
    # TAXABLE INCOME (row 30)
    # =====================
    ws["A30"] = "TAXABLE INCOME"
    ws["A30"].font = Font(bold=True)
    # B30 left empty - agent must add formula

    # =====================
    # TAX CALCULATION (rows 32-40)
    # =====================
    ws["A32"] = "TAX CALCULATION (2024 Brackets - MFJ)"
    ws["A32"].font = Font(bold=True, underline="single")

    # Bracket reference table
    ws["A33"] = "Bracket"
    ws["B33"] = "Rate"
    ws["C33"] = "Upper Limit"
    ws["A33"].font = Font(bold=True)
    ws["B33"].font = Font(bold=True)
    ws["C33"].font = Font(bold=True)
    ws.column_dimensions["C"].width = 16

    brackets = [
        ("10% Bracket", "10%", 23200),
        ("12% Bracket", "12%", 94300),
        ("22% Bracket", "22%", 201050),
        ("24% Bracket", "24%", 383900),
        ("32% Bracket", "32%", 487450),
        ("35% Bracket", "35%", 731200),
        ("37% Bracket", "37%", None),
    ]
    for i, (label, rate, limit) in enumerate(brackets):
        row = 34 + i
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=rate)
        if limit is not None:
            ws.cell(row=row, column=3, value=limit)
        else:
            ws.cell(row=row, column=3, value="No limit")

    # =====================
    # Estimated Tax (row 41) - placeholder label only
    # =====================
    ws["A41"] = "Estimated Federal Tax"
    ws["A41"].font = Font(bold=True)
    # B41 left empty - agent must add formula

    # =====================
    # CREDITS (rows 42-44)
    # =====================
    ws["A42"] = "TAX CREDITS"
    ws["A42"].font = Font(bold=True, underline="single")

    ws["A43"] = "Child Tax Credit (2 children)"
    ws["B43"] = 4000.00

    ws["A44"] = "Education Credit (Lifetime Learning)"
    ws["B44"] = 2000.00

    # =====================
    # WITHHOLDINGS (row 46)
    # =====================
    ws["A46"] = "TOTAL TAX WITHHOLDINGS (W-2)"
    ws["A46"].font = Font(bold=True)
    ws["B46"] = 22350.00

    # =====================
    # RESULT (row 48)
    # =====================
    ws["A48"] = "REFUND (+) / AMOUNT DUE (-)"
    ws["A48"].font = Font(bold=True, size=12)
    # B48 left empty - agent must add formula

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
