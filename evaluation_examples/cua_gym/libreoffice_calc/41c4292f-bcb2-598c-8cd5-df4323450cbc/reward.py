"""
Reward Script: Tax Estimation Worksheet
Task ID: calc_wf_047
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20): AGI formula in B18
  Component 2 (0.10): Total Itemized Deductions formula in B27
  Component 3 (0.10): Deduction Used formula in B28
  Component 4 (0.10): Taxable Income formula in B30
  Component 5 (0.20): Tax bracket calculation formula in B41
  Component 6 (0.10): Refund/Amount Due formula in B48
  Component 7 (0.10): Currency formatting on formula cells
  Component 8 (0.10): Print area set
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_047'


def persist_app_state(domain):
    """Attempt to save any unsaved LibreOffice state."""
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def check_formula_present(ws, coord, required_fragments):
    """
    Check if a cell contains a formula with the required fragments.
    Returns True if cell value is a string starting with '=' and
    contains all required fragments (case-insensitive).
    """
    val = ws[coord].value
    if not isinstance(val, str) or not val.startswith('='):
        return False
    val_upper = val.upper().replace(" ", "")
    for frag in required_fragments:
        if frag.upper().replace(" ", "") not in val_upper:
            return False
    return True


def check_currency_format(ws, coord):
    """Check if a cell has currency number format ($#,##0.00 or similar)."""
    fmt = ws[coord].number_format
    if fmt and '$' in fmt:
        return True
    return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet exists
    if 'Tax Estimate' not in wb.sheetnames:
        print("FAIL: 'Tax Estimate' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Tax Estimate']

    # Component 1: AGI formula in B18 (0.20 points)
    # Should be =SUM(B4:B10)-SUM(B13:B16) or equivalent
    # Initial has NO formula in B18, golden does
    try:
        if check_formula_present(ws, 'B18', ['SUM', 'B4', 'B10', 'B13', 'B16']):
            print(f"PASS: Component 1 -- AGI formula in B18: {ws['B18'].value} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 -- Expected AGI formula in B18, found: {repr(ws['B18'].value)}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Total Itemized Deductions formula in B27 (0.10 points)
    # Should be =SUM(B23:B26) or equivalent
    try:
        if check_formula_present(ws, 'B27', ['SUM', 'B23', 'B26']):
            print(f"PASS: Component 2 -- Itemized deductions formula in B27: {ws['B27'].value} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 -- Expected SUM formula in B27, found: {repr(ws['B27'].value)}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Deduction Used formula in B28 (0.10 points)
    # Should be =MAX(B21,B27) or equivalent
    try:
        if check_formula_present(ws, 'B28', ['MAX', 'B21', 'B27']):
            print(f"PASS: Component 3 -- Deduction used formula in B28: {ws['B28'].value} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 -- Expected MAX formula in B28, found: {repr(ws['B28'].value)}")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Taxable Income formula in B30 (0.10 points)
    # Should be =MAX(B18-B28,0) or equivalent referencing B18 and B28
    try:
        if check_formula_present(ws, 'B30', ['B18', 'B28']):
            print(f"PASS: Component 4 -- Taxable income formula in B30: {ws['B30'].value} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 -- Expected taxable income formula in B30, found: {repr(ws['B30'].value)}")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # Component 5: Tax bracket calculation formula in B41 (0.20 points)
    # Should use nested IF with tax bracket rates (10%, 12%, 22%, etc.)
    # and reference B30 (taxable income)
    try:
        if check_formula_present(ws, 'B41', ['IF', 'B30', '0.1', '0.12', '0.22']):
            print(f"PASS: Component 5 -- Tax bracket formula in B41 (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 5 -- Expected nested IF tax bracket formula in B41, found: {repr(ws['B41'].value)}")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    # Component 6: Refund/Amount Due formula in B48 (0.10 points)
    # Should compute = Withholdings + Credits - Tax = B46 + SUM(B43:B44) - B41 or equivalent
    try:
        if check_formula_present(ws, 'B48', ['B46', 'B41']):
            print(f"PASS: Component 6 -- Refund formula in B48: {ws['B48'].value} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 6 -- Expected refund formula in B48, found: {repr(ws['B48'].value)}")
    except Exception as e:
        print(f"ERROR: Component 6 -- {e}")

    # Component 7: Currency formatting on formula cells (0.10 points)
    # Golden has $#,##0.00 on B4:B10, B13:B16, B18, B21, B23:B26, B27, B28, B30, B41, B43, B44, B46, B48
    # Initial has NO currency formatting (all General)
    # Check a subset of key formula cells that changed
    try:
        currency_cells = ['B18', 'B27', 'B28', 'B30', 'B41', 'B48']
        currency_count = sum(1 for c in currency_cells if check_currency_format(ws, c))
        if currency_count >= 5:
            print(f"PASS: Component 7 -- Currency formatting on {currency_count}/{len(currency_cells)} formula cells (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 7 -- Currency formatting on only {currency_count}/{len(currency_cells)} formula cells")
    except Exception as e:
        print(f"ERROR: Component 7 -- {e}")

    # Component 8: Print area set (0.10 points)
    # Golden has print area 'Tax Estimate'!$A$1:$C$48, initial has none
    try:
        print_area = ws.print_area
        if print_area:
            # print_area can be a string like "'Tax Estimate'!$A$1:$C$48" or a list
            pa_str = str(print_area)
            if 'A' in pa_str and '48' in pa_str:
                print(f"PASS: Component 8 -- Print area set: {pa_str} (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 8 -- Print area set but doesn't cover expected range: {pa_str}")
        else:
            print(f"FAIL: Component 8 -- No print area set")
    except Exception as e:
        print(f"ERROR: Component 8 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
