"""
Initial Setup: Create a Name_Format spreadsheet with ID and First Name columns, no validation.
Task ID: calc_gcv_080
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_080'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Name_Format"

    # --- Headers ---
    headers = ["ID", "First Name"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Data: 24 rows of realistic first names ---
    first_names = [
        "Sarah", "Marcus", "Elena", "David", "Priya",
        "James", "Yuki", "Carlos", "Fatima", "Oliver",
        "Mei", "Alexander", "Zara", "Thomas", "Amara",
        "Benjamin", "Sofia", "Ethan", "Nadia", "Lucas",
        "Ava", "Rafael", "Ingrid", "Samuel",
    ]

    data_font = Font(name="Calibri", size=11)
    data_align = Alignment(horizontal="left", vertical="center")

    for i, name in enumerate(first_names):
        row = i + 2  # rows 2-25

        # Column A: ID
        id_cell = ws.cell(row=row, column=1, value=i + 1)
        id_cell.font = data_font
        id_cell.alignment = Alignment(horizontal="center", vertical="center")
        id_cell.border = thin_border

        # Column B: First Name
        name_cell = ws.cell(row=row, column=2, value=name)
        name_cell.font = data_font
        name_cell.alignment = data_align
        name_cell.border = thin_border

    # --- Column widths ---
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 22

    # --- Freeze header row ---
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f"Initial file created: {OUTPUT}")

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print("GUI_READY: launched LibreOffice Calc with DISPLAY=:0")


create_initial()
