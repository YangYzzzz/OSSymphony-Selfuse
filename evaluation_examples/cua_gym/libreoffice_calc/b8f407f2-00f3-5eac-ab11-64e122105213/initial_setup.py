"""
Initial Setup: Multi-app doc-follow task — create transform_spec.odt and raw_data.ods
Task ID: osworld_multi_apps_doc_follow_instructions_006
Domain: libreoffice_calc (+ libreoffice_writer for spec doc)
"""

import os
import shlex
import subprocess
import time

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_doc_follow_instructions_006'
DOCS_DIR = f'{WORKDIR}/Documents'
DESKTOP_DIR = f'{WORKDIR}/Desktop'
SPEC_FILE = f'{DOCS_DIR}/transform_spec.odt'
DATA_FILE = f'{DESKTOP_DIR}/raw_data.ods'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env['DISPLAY'] = ':0'
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_spec_odt():
    """Create the transform specification ODT document in Documents folder."""
    os.makedirs(DOCS_DIR, exist_ok=True)

    # Use python-docx to generate content, then save as docx,
    # but since we need ODT format we'll use odfpy
    try:
        from odf.opendocument import OpenDocumentText
        from odf.style import Style, TextProperties, ParagraphProperties
        from odf.text import H, P, List, ListItem
        from odf.namespaces import TEXTNS

        doc = OpenDocumentText()

        # Create heading style
        heading_style = Style(name='Heading1', family='paragraph')
        heading_style.addElement(TextProperties(fontsize='16pt', fontweight='bold'))
        doc.automaticstyles.addElement(heading_style)

        # Create body style
        body_style = Style(name='BodyText', family='paragraph')
        body_style.addElement(TextProperties(fontsize='12pt'))
        doc.automaticstyles.addElement(body_style)

        # Title heading
        h1 = H(outlinelevel=1, text='Data Transformation Specification')
        doc.text.addElement(h1)

        # Introduction
        intro = P(text='Apply the following 7 transformations to the file raw_data.ods on the Desktop:')
        doc.text.addElement(intro)

        # Blank line
        doc.text.addElement(P(text=''))

        # Requirements as numbered items
        requirements = [
            'Requirement 1: Add column H named "Profit_Margin". Formula: =(Revenue-Cost)/Revenue, formatted as percentage (0.00%).',
            'Requirement 2: Add column I named "Performance". Formula: =IF(Profit_Margin>0.3,"High",IF(Profit_Margin>0.15,"Medium","Low")) referencing column H.',
            'Requirement 3: Apply conditional formatting to all data cells in column I. Use green background (#00FF00) for "High", yellow background (#FFFF00) for "Medium", red background (#FF0000) for "Low".',
            'Requirement 4: Sort the data rows (excluding header and any summary rows) by Revenue column in descending order.',
            'Requirement 5: Insert a SUBTOTAL row at row 2 (immediately after the header row). This row should show the total Revenue and total Cost using SUM formulas.',
            'Requirement 6: Freeze the first two rows (header row and subtotal row) so they remain visible when scrolling.',
            'Requirement 7: Hide column A (the internal ID column) from view.',
        ]

        for req in requirements:
            p = P(text=req)
            doc.text.addElement(p)
            doc.text.addElement(P(text=''))

        # Footer note
        doc.text.addElement(P(text='Note: Save the file as raw_data.ods after applying all transformations.'))

        doc.save(SPEC_FILE)
        print(f'Spec ODT created: {SPEC_FILE}')

    except ImportError:
        # Fallback: create a simple text-based ODT using zipfile
        import zipfile
        import io

        content_xml = '''<?xml version="1.0" encoding="UTF-8"?>
<office:document-content
  xmlns:office="urn:oasis:names:tc:opendocument:xmlns:office:1.0"
  xmlns:text="urn:oasis:names:tc:opendocument:xmlns:text:1.0"
  xmlns:style="urn:oasis:names:tc:opendocument:xmlns:style:1.0"
  office:version="1.2">
  <office:body>
    <office:text>
      <text:h text:outline-level="1">Data Transformation Specification</text:h>
      <text:p>Apply the following 7 transformations to the file raw_data.ods on the Desktop:</text:p>
      <text:p></text:p>
      <text:p>Requirement 1: Add column H named "Profit_Margin". Formula: =(Revenue-Cost)/Revenue, formatted as percentage (0.00%).</text:p>
      <text:p>Requirement 2: Add column I named "Performance". Formula: =IF(Profit_Margin&gt;0.3,"High",IF(Profit_Margin&gt;0.15,"Medium","Low")) referencing column H.</text:p>
      <text:p>Requirement 3: Apply conditional formatting to all data cells in column I. Use green background (#00FF00) for "High", yellow background (#FFFF00) for "Medium", red background (#FF0000) for "Low".</text:p>
      <text:p>Requirement 4: Sort the data rows (excluding header and any summary rows) by Revenue column in descending order.</text:p>
      <text:p>Requirement 5: Insert a SUBTOTAL row at row 2 (immediately after the header row). This row should show the total Revenue and total Cost using SUM formulas.</text:p>
      <text:p>Requirement 6: Freeze the first two rows (header row and subtotal row) so they remain visible when scrolling.</text:p>
      <text:p>Requirement 7: Hide column A (the internal ID column) from view.</text:p>
      <text:p></text:p>
      <text:p>Note: Save the file as raw_data.ods after applying all transformations.</text:p>
    </office:text>
  </office:body>
</office:document-content>'''

        manifest_xml = '''<?xml version="1.0" encoding="UTF-8"?>
<manifest:manifest xmlns:manifest="urn:oasis:names:tc:opendocument:xmlns:manifest:1.0">
  <manifest:file-entry manifest:full-path="/" manifest:media-type="application/vnd.oasis.opendocument.text"/>
  <manifest:file-entry manifest:full-path="content.xml" manifest:media-type="text/xml"/>
  <manifest:file-entry manifest:full-path="meta.xml" manifest:media-type="text/xml"/>
</manifest:manifest>'''

        meta_xml = '''<?xml version="1.0" encoding="UTF-8"?>
<office:document-meta xmlns:office="urn:oasis:names:tc:opendocument:xmlns:office:1.0"
                      office:version="1.2">
  <office:meta/>
</office:document-meta>'''

        mimetype = 'application/vnd.oasis.opendocument.text'

        with zipfile.ZipFile(SPEC_FILE, 'w', zipfile.ZIP_DEFLATED) as zf:
            # mimetype must be first and uncompressed
            zf.writestr(zipfile.ZipInfo('mimetype'), mimetype, compress_type=zipfile.ZIP_STORED)
            zf.writestr('content.xml', content_xml)
            zf.writestr('META-INF/manifest.xml', manifest_xml)
            zf.writestr('meta.xml', meta_xml)

        print(f'Spec ODT created (fallback zip method): {SPEC_FILE}')


def create_raw_data_ods():
    """Create raw_data.ods on the Desktop with 25 rows, columns A-G."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    os.makedirs(DESKTOP_DIR, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sales Data'

    # Column headers: A=ID (internal), B=Product, C=Category, D=Revenue, E=Cost, F=Units, G=Quarter
    headers = ['ID', 'Product', 'Category', 'Revenue', 'Cost', 'Units', 'Quarter']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # 25 rows of realistic business data
    # Data will NOT be sorted by Revenue (raw order) so agent has sorting to do
    data = [
        # ID, Product, Category, Revenue, Cost, Units, Quarter
        [1001, 'CloudSync Pro', 'Software', 128500.00, 42300.00, 215, 'Q1-2025'],
        [1002, 'DataVault Basic', 'Storage', 43200.00, 18700.00, 432, 'Q1-2025'],
        [1003, 'NetGuard Enterprise', 'Security', 287400.00, 104200.00, 89, 'Q1-2025'],
        [1004, 'OfficeMax Suite', 'Productivity', 95600.00, 31800.00, 356, 'Q2-2025'],
        [1005, 'StreamLine Analytics', 'Analytics', 176300.00, 83500.00, 127, 'Q2-2025'],
        [1006, 'SecureVPN Plus', 'Security', 62800.00, 21400.00, 511, 'Q2-2025'],
        [1007, 'DevTools Pro', 'Development', 39700.00, 15200.00, 198, 'Q2-2025'],
        [1008, 'CloudSync Basic', 'Software', 58900.00, 22600.00, 589, 'Q3-2025'],
        [1009, 'DataVault Pro', 'Storage', 112400.00, 47800.00, 298, 'Q3-2025'],
        [1010, 'MailGuard Suite', 'Security', 74300.00, 28900.00, 421, 'Q3-2025'],
        [1011, 'BizAnalytics Premium', 'Analytics', 215700.00, 97300.00, 143, 'Q3-2025'],
        [1012, 'ProjectMaster Pro', 'Productivity', 88100.00, 34200.00, 267, 'Q4-2025'],
        [1013, 'NetWatch Lite', 'Security', 31500.00, 14800.00, 630, 'Q4-2025'],
        [1014, 'CloudHosting Plus', 'Infrastructure', 193800.00, 88600.00, 94, 'Q4-2025'],
        [1015, 'DevStudio Enterprise', 'Development', 342600.00, 128900.00, 76, 'Q4-2025'],
        [1016, 'TaskSync Basic', 'Productivity', 27400.00, 11900.00, 548, 'Q1-2025'],
        [1017, 'DataStream API', 'Analytics', 156900.00, 71200.00, 183, 'Q1-2025'],
        [1018, 'SecureSign Pro', 'Security', 48200.00, 19700.00, 362, 'Q2-2025'],
        [1019, 'MobileFirst Platform', 'Development', 124700.00, 53400.00, 211, 'Q2-2025'],
        [1020, 'CloudBackup Pro', 'Storage', 81300.00, 34100.00, 325, 'Q3-2025'],
        [1021, 'ReportBuilder Plus', 'Analytics', 67500.00, 29800.00, 450, 'Q3-2025'],
        [1022, 'TeamCollab Suite', 'Productivity', 103400.00, 43600.00, 288, 'Q4-2025'],
        [1023, 'NetMonitor Pro', 'Infrastructure', 145200.00, 67300.00, 196, 'Q4-2025'],
        [1024, 'APIGateway Enterprise', 'Development', 268900.00, 112400.00, 87, 'Q1-2025'],
        [1025, 'DataWarehouse Basic', 'Storage', 54700.00, 24300.00, 437, 'Q2-2025'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            # Format currency columns (Revenue=D=col4, Cost=E=col5)
            if c in (4, 5):
                cell.number_format = '#,##0.00'

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 8   # ID
    ws.column_dimensions['B'].width = 25  # Product
    ws.column_dimensions['C'].width = 16  # Category
    ws.column_dimensions['D'].width = 14  # Revenue
    ws.column_dimensions['E'].width = 14  # Cost
    ws.column_dimensions['F'].width = 10  # Units
    ws.column_dimensions['G'].width = 12  # Quarter

    # Save as xlsx first (openpyxl doesn't support ods directly),
    # then we'll use the xlsx file but name it .ods for the task
    # Actually save as xlsx and rename — or just save as xlsx with .ods extension
    # LibreOffice will open either format
    # For compatibility, save as .xlsx but task refers to .ods
    # We'll save to a temp xlsx, then use subprocess to convert to ods
    import tempfile
    tmp_xlsx = f'{DESKTOP_DIR}/raw_data_tmp.xlsx'
    wb.save(tmp_xlsx)
    print(f'Temporary XLSX created: {tmp_xlsx}')

    # Convert to ODS format using LibreOffice headless
    try:
        env = os.environ.copy()
        env['DISPLAY'] = ':0'
        result = subprocess.run(
            ['libreoffice', '--headless', '--convert-to', 'ods',
             '--outdir', DESKTOP_DIR, tmp_xlsx],
            capture_output=True, text=True, timeout=60, env=env
        )
        print(f'LibreOffice conversion stdout: {result.stdout}')
        print(f'LibreOffice conversion stderr: {result.stderr}')

        # The converted file will be raw_data_tmp.ods, rename to raw_data.ods
        converted = f'{DESKTOP_DIR}/raw_data_tmp.ods'
        if os.path.exists(converted):
            if os.path.exists(DATA_FILE):
                os.remove(DATA_FILE)
            os.rename(converted, DATA_FILE)
            print(f'Renamed to: {DATA_FILE}')
        else:
            # If conversion failed, just keep xlsx with .ods name
            import shutil
            shutil.copy(tmp_xlsx, DATA_FILE)
            print(f'Conversion failed, copied xlsx as: {DATA_FILE}')
    except Exception as e:
        print(f'Conversion error: {e}, copying xlsx as ods')
        import shutil
        shutil.copy(tmp_xlsx, DATA_FILE)

    # Clean up temp file
    if os.path.exists(tmp_xlsx):
        os.remove(tmp_xlsx)

    print(f'Raw data file created: {DATA_FILE}')


def main():
    print('=== Initial Setup: osworld_multi_apps_doc_follow_instructions_006 ===')

    # Create the spec ODT document
    create_spec_odt()

    # Create the raw data ODS file on Desktop
    create_raw_data_ods()

    # Verify files exist
    if not os.path.exists(SPEC_FILE):
        raise FileNotFoundError(f'Spec file not created: {SPEC_FILE}')
    if not os.path.exists(DATA_FILE):
        raise FileNotFoundError(f'Data file not created: {DATA_FILE}')

    print(f'All files created successfully.')
    print(f'  Spec: {SPEC_FILE}')
    print(f'  Data: {DATA_FILE}')

    # GUI startup: open both files for the agent
    # First open the spec document in Writer so agent can read requirements
    time.sleep(1.0)
    launch_gui(f'libreoffice --writer "{SPEC_FILE}"', delay_sec=2.5)

    # Then open the data file in Calc
    launch_gui(f'libreoffice --calc "{DATA_FILE}"', delay_sec=2.5)

    print('GUI_READY: launched LibreOffice Writer (spec) and Calc (data) with DISPLAY=:0')


main()
