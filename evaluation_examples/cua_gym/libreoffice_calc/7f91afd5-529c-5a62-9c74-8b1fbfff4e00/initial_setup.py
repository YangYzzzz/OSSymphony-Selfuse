"""
Initial Setup: Apply conditional formatting for overdue tasks
Task ID: calc_gfl_088
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
from datetime import date, timedelta
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_088'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Tasks'

    # Headers
    headers = ['Task ID', 'Name', 'Assignee', 'Due Date', 'Priority', 'Status', 'Notes']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 34 task records (rows 2-35)
    today = date.today()

    tasks = [
        ['T-001', 'Redesign landing page', 'Sarah Chen', today - timedelta(days=45), 'High', 'In Progress', 'Waiting for brand assets from marketing team'],
        ['T-002', 'Fix login timeout bug', 'Marcus Johnson', today - timedelta(days=30), 'Critical', 'Open', 'Users report session expires after 5 minutes'],
        ['T-003', 'Update API documentation', 'Elena Rodriguez', today + timedelta(days=14), 'Medium', 'Open', 'Swagger docs need refresh for v2.3 endpoints'],
        ['T-004', 'Database migration script', 'James Liu', today - timedelta(days=12), 'High', 'In Progress', 'Migrating from PostgreSQL 14 to 16'],
        ['T-005', 'Customer onboarding flow', 'Priya Sharma', today + timedelta(days=30), 'High', 'Open', 'New wizard-based onboarding for enterprise clients'],
        ['T-006', 'Performance audit Q1', 'David Kim', today - timedelta(days=60), 'Medium', 'Completed', 'Lighthouse scores improved by 15 points'],
        ['T-007', 'Mobile app push notifications', 'Aisha Patel', today + timedelta(days=7), 'High', 'In Progress', 'Integrate Firebase Cloud Messaging'],
        ['T-008', 'Security vulnerability patch', 'Tom Nakamura', today - timedelta(days=5), 'Critical', 'Open', 'CVE-2026-1234 affects auth middleware'],
        ['T-009', 'A/B testing framework setup', 'Laura Martinez', today + timedelta(days=21), 'Medium', 'Open', 'Evaluate Optimizely vs custom solution'],
        ['T-010', 'Quarterly sales report', 'Robert Chang', today - timedelta(days=3), 'High', 'In Progress', 'Dashboard for executive review meeting'],
        ['T-011', 'Refactor payment module', 'Nina Kowalski', today + timedelta(days=45), 'Medium', 'Open', 'Extract Stripe logic into separate service'],
        ['T-012', 'User feedback survey', 'Chris Thompson', today - timedelta(days=20), 'Low', 'Completed', 'NPS survey sent to 5000 active users'],
        ['T-013', 'CI/CD pipeline optimization', 'Yuki Tanaka', today + timedelta(days=10), 'High', 'In Progress', 'Build time currently 18 min, target under 8 min'],
        ['T-014', 'GDPR compliance review', 'Sophie Laurent', today - timedelta(days=8), 'Critical', 'Open', 'Annual audit for EU data processing requirements'],
        ['T-015', 'Design system color tokens', 'Miguel Santos', today + timedelta(days=60), 'Low', 'Open', 'Align with new brand guidelines from Q4'],
        ['T-016', 'Load testing infrastructure', 'Rachel Green', today - timedelta(days=15), 'High', 'In Progress', 'Set up k6 scripts for checkout flow'],
        ['T-017', 'Email template refresh', 'Alex Okonkwo', today + timedelta(days=5), 'Medium', 'Open', 'Update transactional emails to new brand style'],
        ['T-018', 'Inventory sync module', 'Hannah Lee', today - timedelta(days=25), 'High', 'Open', 'Real-time sync between warehouse and e-commerce'],
        ['T-019', 'SSO integration', 'Daniel Fischer', today + timedelta(days=35), 'High', 'Open', 'Support SAML 2.0 for enterprise customers'],
        ['T-020', 'Bug: Cart total rounding', 'Maria Gonzalez', today - timedelta(days=2), 'Medium', 'In Progress', 'Cent-level rounding errors on multi-item orders'],
        ['T-021', 'Accessibility audit', 'Kevin Park', today + timedelta(days=18), 'High', 'Open', 'WCAG 2.1 AA compliance check for main flows'],
        ['T-022', 'Analytics dashboard v2', 'Fatima Al-Rashid', today - timedelta(days=40), 'Medium', 'Completed', 'Added funnel visualization and cohort analysis'],
        ['T-023', 'Chatbot training data', 'Lucas Weber', today + timedelta(days=25), 'Low', 'Open', 'Curate 2000 QA pairs for support chatbot'],
        ['T-024', 'Rate limiter implementation', 'Olivia Nguyen', today - timedelta(days=7), 'High', 'Open', 'Protect API endpoints from abuse, 100 req/min'],
        ['T-025', 'Vendor portal prototype', 'Benjamin Muller', today + timedelta(days=50), 'Medium', 'Open', 'Self-service portal for supplier management'],
        ['T-026', 'Log aggregation setup', 'Chloe Dubois', today - timedelta(days=18), 'High', 'In Progress', 'Centralize logs with ELK stack'],
        ['T-027', 'Product image CDN migration', 'Isaac Yamamoto', today + timedelta(days=12), 'Medium', 'Open', 'Move from S3 direct to CloudFront distribution'],
        ['T-028', 'Customer churn prediction', 'Amara Osei', today - timedelta(days=35), 'Medium', 'Open', 'ML model to identify at-risk accounts'],
        ['T-029', 'Multi-language support', 'Viktor Petrov', today + timedelta(days=40), 'High', 'Open', 'i18n for French, German, Spanish, Japanese'],
        ['T-030', 'Backup verification script', 'Zara Hussain', today - timedelta(days=10), 'Critical', 'In Progress', 'Automated restore testing for DR compliance'],
        ['T-031', 'Webhook retry mechanism', 'Nathan Brooks', today + timedelta(days=8), 'Medium', 'Open', 'Exponential backoff for failed webhook deliveries'],
        ['T-032', 'Dark mode implementation', 'Emily Watson', today - timedelta(days=1), 'Low', 'Open', 'CSS variable-based theme switching'],
        ['T-033', 'API versioning strategy', 'Raj Kapoor', today + timedelta(days=20), 'High', 'Open', 'Define v3 migration path and deprecation timeline'],
        ['T-034', 'Search indexing optimization', 'Lena Bergstrom', today - timedelta(days=22), 'High', 'In Progress', 'Elasticsearch reindex taking too long on catalog'],
    ]

    for r, row_data in enumerate(tasks, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 50

    # Format date column
    for r in range(2, 36):
        ws.cell(row=r, column=4).number_format = 'yyyy-mm-dd'

    # NO conditional formatting in initial state
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
