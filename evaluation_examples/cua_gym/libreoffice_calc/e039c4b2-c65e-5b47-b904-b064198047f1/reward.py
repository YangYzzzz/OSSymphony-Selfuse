"""
Reward Script: Select sheets Q1, Q2, and Q3 simultaneously for grouped printing
Task ID: calc_sht_multisel_001
Domain: libreoffice_calc
Scoring:
  - Component 1: Sheet 'Q2' is selected (tabSelected=True)  — 0.4 pts
  - Component 2: Sheet 'Q3' is selected (tabSelected=True)  — 0.4 pts
  - Component 3: Sheets 'Annual' and 'Q4' are NOT selected  — 0.2 pts
  Total: 1.0

Rationale:
  The task requires the agent to group three sheets (Q1, Q2, Q3) together for printing.
  In the initial state, only Q1 is selected (tabSelected=True). The task is complete when
  Q2 and Q3 are also selected (tabSelected=True), making all three sheets grouped.
  'Annual' and 'Q4' must remain unselected.

  We do NOT score Q1's selection state because Q1 is already selected in the initial file.
  Scoring Q1 would credit a pre-existing precondition, not a task-introduced change.
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_sht_multisel_001'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook — if this fails, nothing can be verified
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: verify the expected sheets exist
    expected_sheets = {'Annual', 'Q1', 'Q2', 'Q3', 'Q4'}
    actual_sheets = set(wb.sheetnames)
    if not expected_sheets.issubset(actual_sheets):
        missing = expected_sheets - actual_sheets
        print(f"CRITICAL: Expected sheets missing: {missing}. Workbook has: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Sheet 'Q2' is selected/grouped (tabSelected=True) — 0.4 points
    # In the initial file, Q2 has tabSelected=False.
    # This only passes after the agent selects Q2 as part of the group.
    try:
        ws_q2 = wb['Q2']
        q2_selected = ws_q2.sheet_view.tabSelected
        if q2_selected:
            print(f"PASS: Component 1 — Sheet 'Q2' is selected (tabSelected=True) (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — Sheet 'Q2' is NOT selected (tabSelected={q2_selected}), expected True")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not check Q2 tabSelected: {e}")

    # Component 2: Sheet 'Q3' is selected/grouped (tabSelected=True) — 0.4 points
    # In the initial file, Q3 has tabSelected=False.
    # This only passes after the agent selects Q3 as part of the group.
    try:
        ws_q3 = wb['Q3']
        q3_selected = ws_q3.sheet_view.tabSelected
        if q3_selected:
            print(f"PASS: Component 2 — Sheet 'Q3' is selected (tabSelected=True) (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 2 — Sheet 'Q3' is NOT selected (tabSelected={q3_selected}), expected True")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not check Q3 tabSelected: {e}")

    # Component 3: 'Annual' and 'Q4' are NOT selected — 0.2 points
    # This ensures the agent did not over-select (e.g., select all sheets).
    # In the initial file, Annual=False and Q4=False, but this component is compound:
    # it only awards points WHEN combined with components 1 and 2 passing (i.e., the group
    # must exist AND be correctly bounded). As a standalone gating check it guards against
    # the degenerate case where the agent selects all 5 sheets.
    # NOTE: Because Annual and Q4 are also False in the initial file, this component alone
    # would pass on the initial file. Therefore we add a guard: this component only awards
    # points if at least one of Q2 or Q3 is selected (i.e., the group was actually formed).
    try:
        ws_annual = wb['Annual']
        ws_q4 = wb['Q4']
        annual_not_selected = not ws_annual.sheet_view.tabSelected
        q4_not_selected = not ws_q4.sheet_view.tabSelected

        # Guard: only award scope-correctness points if the group was actually formed
        group_formed = (wb['Q2'].sheet_view.tabSelected or wb['Q3'].sheet_view.tabSelected)

        if annual_not_selected and q4_not_selected and group_formed:
            print(f"PASS: Component 3 — 'Annual' and 'Q4' are correctly excluded from selection (0.2 pts)")
            total_score += 0.2
        elif not annual_not_selected:
            print(f"FAIL: Component 3 — Sheet 'Annual' should NOT be selected, but tabSelected={ws_annual.sheet_view.tabSelected}")
        elif not q4_not_selected:
            print(f"FAIL: Component 3 — Sheet 'Q4' should NOT be selected, but tabSelected={ws_q4.sheet_view.tabSelected}")
        else:
            print(f"FAIL: Component 3 — Group not formed yet (Q2={wb['Q2'].sheet_view.tabSelected}, Q3={wb['Q3'].sheet_view.tabSelected})")
    except Exception as e:
        print(f"ERROR: Component 3 — Could not check Annual/Q4 tabSelected: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
