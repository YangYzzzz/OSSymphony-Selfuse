"""
Initial Setup: Workbook with 5 sheets for quarterly P&L reporting.
Task ID: calc_sht_multisel_001
Domain: libreoffice_calc

Creates a workbook with Annual, Q1, Q2, Q3, Q4 sheets.
Only Q1 is selected/active (tabSelected=True).
Q2 and Q3 are NOT selected — the user must group them.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_sht_multisel_001'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

# P&L row labels (rows 2-50 for data, row 1 is header)
PL_CATEGORIES = [
    # Revenue section
    'Product Sales',
    'Service Revenue',
    'Subscription Revenue',
    'License Fees',
    'Other Revenue',
    'Total Revenue',
    '',
    # Cost of Goods Sold
    'Material Costs',
    'Labor Costs',
    'Manufacturing Overhead',
    'Total COGS',
    '',
    # Gross Profit
    'Gross Profit',
    '',
    # Operating Expenses
    'Salaries & Wages',
    'Rent & Utilities',
    'Marketing & Advertising',
    'Research & Development',
    'Depreciation',
    'Administrative Expenses',
    'Travel & Entertainment',
    'Insurance',
    'Total Operating Expenses',
    '',
    # Operating Income
    'Operating Income (EBIT)',
    '',
    # Other Income/Expenses
    'Interest Income',
    'Interest Expense',
    'Other Income',
    'Other Expense',
    'Total Other Income/(Expense)',
    '',
    # Net Income
    'Income Before Tax',
    'Income Tax Expense',
    'Net Income',
    '',
    # Additional metrics
    'EBITDA',
    'Gross Margin %',
    'Operating Margin %',
    'Net Margin %',
]

# Quarterly data: [Q1, Q2, Q3, Q4] values for each category
# Realistic P&L figures for a mid-size company (in thousands USD)
QUARTERLY_DATA = {
    'Product Sales':              [1_245_800, 1_318_400, 1_402_700, 1_589_300],
    'Service Revenue':            [  387_500,   412_000,   445_600,   498_200],
    'Subscription Revenue':       [  156_200,   163_800,   172_400,   181_600],
    'License Fees':               [   42_300,    38_700,    51_200,    67_800],
    'Other Revenue':              [   18_500,    14_200,    22_100,    29_400],
    'Total Revenue':              [1_850_300, 1_947_100, 2_094_000, 2_366_300],
    '': None,
    'Material Costs':             [  487_200,   513_400,   551_800,   621_900],
    'Labor Costs':                [  312_500,   328_700,   347_200,   389_600],
    'Manufacturing Overhead':     [   98_400,   102_100,   108_700,   124_300],
    'Total COGS':                 [  898_100,   944_200, 1_007_700, 1_135_800],
    'Gross Profit':               [  952_200, 1_002_900, 1_086_300, 1_230_500],
    'Salaries & Wages':           [  298_400,   312_700,   325_100,   341_800],
    'Rent & Utilities':           [   42_600,    43_200,    43_800,    44_500],
    'Marketing & Advertising':    [   87_300,    94_600,   108_400,   142_700],
    'Research & Development':     [   65_800,    71_200,    73_400,    78_100],
    'Depreciation':               [   24_500,    24_500,    24_500,    24_500],
    'Administrative Expenses':    [   38_700,    41_200,    43_800,    47_300],
    'Travel & Entertainment':     [   12_400,    11_800,    13_200,    18_700],
    'Insurance':                  [    8_900,     8_900,     9_200,     9_200],
    'Total Operating Expenses':   [  578_600,   608_100,   641_400,   706_800],
    'Operating Income (EBIT)':    [  373_600,   394_800,   444_900,   523_700],
    'Interest Income':            [    5_200,     5_800,     6_100,     6_400],
    'Interest Expense':           [  -18_700,   -18_700,   -17_900,   -17_200],
    'Other Income':               [    3_800,     2_400,     7_200,     4_100],
    'Other Expense':              [   -2_100,    -1_800,    -3_400,    -2_700],
    'Total Other Income/(Expense)': [-11_800, -12_300,    -8_000,    -9_400],
    'Income Before Tax':          [  361_800,   382_500,   436_900,   514_300],
    'Income Tax Expense':         [   86_832,    91_800,   104_856,   123_432],
    'Net Income':                 [  274_968,   290_700,   332_044,   390_868],
    'EBITDA':                     [  398_100,   419_300,   469_400,   548_200],
    'Gross Margin %':             [    51.46,     51.51,     51.88,     51.99],
    'Operating Margin %':         [    20.19,     20.28,     21.24,     22.13],
    'Net Margin %':               [    14.86,     14.93,     15.85,     16.52],
}

QUARTER_HEADERS = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
# Column mapping: A=Category, B=Jan, C=Feb, D=Mar, E=Q Total, F=Budget, G=Variance, H=Var%

MONTHS_BY_QUARTER = {
    'Q1': ['January', 'February', 'March'],
    'Q2': ['April', 'May', 'June'],
    'Q3': ['July', 'August', 'September'],
    'Q4': ['October', 'November', 'December'],
}

QUARTER_MONTH_SPLIT = {
    # Q index -> [month1_pct, month2_pct, month3_pct]
    0: [0.31, 0.33, 0.36],  # Q1
    1: [0.32, 0.34, 0.34],  # Q2
    2: [0.33, 0.33, 0.34],  # Q3
    3: [0.30, 0.33, 0.37],  # Q4
}

header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
header_fill = PatternFill(start_color='FF2E4057', end_color='FF2E4057', fill_type='solid')
section_fill = PatternFill(start_color='FFBDD7EE', end_color='FFBDD7EE', fill_type='solid')
total_fill = PatternFill(start_color='FFDDEBF7', end_color='FFDDEBF7', fill_type='solid')
section_font = Font(name='Calibri', size=10, bold=True)
data_font = Font(name='Calibri', size=10)
thin_border_side = Side(style='thin', color='FFB8B8B8')
thin_border = Border(
    left=thin_border_side, right=thin_border_side,
    top=thin_border_side, bottom=thin_border_side
)

SECTION_TOTALS = {
    'Total Revenue', 'Total COGS', 'Gross Profit',
    'Total Operating Expenses', 'Operating Income (EBIT)',
    'Total Other Income/(Expense)', 'Income Before Tax',
    'Income Tax Expense', 'Net Income', 'EBITDA',
    'Gross Margin %', 'Operating Margin %', 'Net Margin %',
}


def add_quarter_sheet(wb, q_name, q_idx):
    """Create a quarterly P&L sheet."""
    ws = wb.create_sheet(q_name)
    months = MONTHS_BY_QUARTER[q_name]
    splits = QUARTER_MONTH_SPLIT[q_idx]

    # Row 1: Headers
    headers = ['Category', months[0], months[1], months[2], f'{q_name} Total', 'Budget', 'Variance', 'Var%']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[1].height = 22

    # Column widths
    col_widths = [28, 14, 14, 14, 14, 14, 14, 10]
    for i, width in enumerate(col_widths, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = width

    # Data rows
    row = 2
    for cat in PL_CATEGORIES:
        if cat == '':
            ws.row_dimensions[row].height = 8
            row += 1
            continue

        q_val = QUARTERLY_DATA.get(cat)
        if q_val is None:
            row += 1
            continue

        total = q_val[q_idx]
        m1 = round(total * splits[0])
        m2 = round(total * splits[1])
        m3 = total - m1 - m2

        # Budget: slight variation from actual
        budget = round(total * (0.97 + (hash(cat + q_name) % 7) * 0.01))
        variance = total - budget
        var_pct = round(variance / budget * 100, 2) if budget != 0 else 0

        is_pct = cat in {'Gross Margin %', 'Operating Margin %', 'Net Margin %}'}
        is_total = cat in SECTION_TOTALS

        values = [cat, m1, m2, m3, total, budget, variance, var_pct]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.font = section_font if is_total else data_font
            if is_total:
                cell.fill = total_fill
            if col == 1:
                cell.alignment = Alignment(horizontal='left', indent=1 if not is_total else 0)
            elif col == 8:
                cell.number_format = '0.00%' if is_pct else '+0.00%;-0.00%'
                cell.alignment = Alignment(horizontal='right')
            elif col in (2, 3, 4, 5, 6, 7):
                if cat in {'Gross Margin %', 'Operating Margin %', 'Net Margin %'}:
                    cell.number_format = '0.00'
                else:
                    cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
        row += 1

    # Freeze header row
    ws.freeze_panes = 'B2'

    # Company and period info in merged cells above header
    # (No actual data merge here to keep it simple)
    return ws


def add_annual_sheet(wb):
    """Create the Annual aggregation sheet."""
    ws = wb.active
    ws.title = 'Annual'

    # Headers
    headers = ['Category', 'Q1', 'Q2', 'Q3', 'Q4', 'Annual Total', 'Budget', 'Variance']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[1].height = 22

    # Column widths
    col_widths = [28, 14, 14, 14, 14, 15, 14, 14]
    for i, w in enumerate(col_widths, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 2
    for cat in PL_CATEGORIES:
        if cat == '':
            ws.row_dimensions[row].height = 8
            row += 1
            continue
        q_val = QUARTERLY_DATA.get(cat)
        if q_val is None:
            row += 1
            continue

        annual_total = sum(q_val)
        budget = round(annual_total * (0.97 + (hash(cat + 'Annual') % 7) * 0.01))
        variance = annual_total - budget
        is_total = cat in SECTION_TOTALS

        values = [cat, q_val[0], q_val[1], q_val[2], q_val[3], annual_total, budget, variance]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.font = section_font if is_total else data_font
            if is_total:
                cell.fill = total_fill
            if col == 1:
                cell.alignment = Alignment(horizontal='left', indent=1 if not is_total else 0)
            elif col >= 2:
                if cat in {'Gross Margin %', 'Operating Margin %', 'Net Margin %'}:
                    cell.number_format = '0.00'
                else:
                    cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
        row += 1

    ws.freeze_panes = 'B2'
    return ws


def create_initial():
    wb = openpyxl.Workbook()

    # Sheet 1: Annual (active sheet by default from Workbook())
    add_annual_sheet(wb)

    # Sheets 2-5: Quarterly P&L
    for i, q in enumerate(['Q1', 'Q2', 'Q3', 'Q4']):
        add_quarter_sheet(wb, q, i)

    # Set ONLY Q1 as the active/selected tab
    # In openpyxl, tabSelected is set via sheet_view
    for ws in wb.worksheets:
        # Remove any tabSelected by setting it to 0 for all sheets
        for sv in ws.sheet_view.sheetViews if hasattr(ws.sheet_view, 'sheetViews') else [ws.sheet_view]:
            sv.tabSelected = False

    # Activate Q1 only
    wb['Q1'].sheet_view.tabSelected = True
    wb.active = wb['Q1']

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheets: {wb.sheetnames}')
    print(f'Active sheet: {wb.active.title}')
    print(f'Tab selected states:')
    for ws in wb.worksheets:
        print(f'  {ws.title}: tabSelected={ws.sheet_view.tabSelected}')


create_initial()
