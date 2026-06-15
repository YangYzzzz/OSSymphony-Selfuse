"""
Reward Script: Invoice generator with line totals, subtotal, tax lookup, tax amount, shipping, and grand total.
Task ID: calc_sales_071
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Line total formulas in D6:D9 (=Bn*Cn)
  Component 2 (0.20): Subtotal formula in B11 (=SUM(D6:D9))
  Component 3 (0.20): Tax rate lookup formula in B12 (VLOOKUP of state from TaxRates)
  Component 4 (0.15): Tax amount formula in B13 (=B11*B12)
  Component 5 (0.15): Grand total formula in B15 (=B11+B13+B14)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_071'


def normalize_formula(val):
    """Normalize a formula string for comparison: uppercase, strip spaces."""
    if not isinstance(val, str):
        return ""
    return val.upper().replace(" ", "")


def is_multiply_formula(val, row):
    """Check if val is a formula that multiplies B<row>*C<row> (in any order)."""
    norm = normalize_formula(val)
    if not norm.startswith("="):
        return False
    expr = norm[1:]  # strip leading '='
    b_ref = f"B{row}"
    c_ref = f"C{row}"
    # Accept =B*C or =C*B
    if expr == f"{b_ref}*{c_ref}" or expr == f"{c_ref}*{b_ref}":
        return True
    return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Invoice sheet must exist
    if 'Invoice' not in wb.sheetnames:
        print("CRITICAL: 'Invoice' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Invoice']

    # Component 1: Line total formulas D6:D9 (0.30 points — 0.075 each)
    # Each cell should contain =Bn*Cn where n is the row number
    try:
        line_total_score = 0.0
        for row in [6, 7, 8, 9]:
            cell_val = ws.cell(row=row, column=4).value  # Column D
            if is_multiply_formula(cell_val, row):
                print(f"PASS: D{row} has multiplication formula: {cell_val} (0.075 pts)")
                line_total_score += 0.075
            else:
                print(f"FAIL: D{row} expected =B{row}*C{row}, found: {repr(cell_val)}")
        if line_total_score > 0:
            total_score += line_total_score
    except Exception as e:
        print(f"ERROR: Component 1 (line totals) — {e}")

    # Component 2: Subtotal formula in B11 (0.20 points)
    # Should be =SUM(D6:D9)
    try:
        b11_val = ws['B11'].value
        b11_norm = normalize_formula(b11_val)
        if b11_norm == "=SUM(D6:D9)":
            print(f"PASS: B11 has SUM formula: {b11_val} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: B11 expected =SUM(D6:D9), found: {repr(b11_val)}")
    except Exception as e:
        print(f"ERROR: Component 2 (subtotal) — {e}")

    # Component 3: Tax rate lookup formula in B12 (0.20 points)
    # Should be a VLOOKUP referencing TaxRates sheet for "CA" state
    try:
        b12_val = ws['B12'].value
        b12_norm = normalize_formula(b12_val)
        # Accept variations: VLOOKUP("CA",...) or VLOOKUP(reference_to_state,...)
        if "VLOOKUP" in b12_norm and "TAXRATES" in b12_norm:
            print(f"PASS: B12 has VLOOKUP formula referencing TaxRates: {b12_val} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: B12 expected VLOOKUP with TaxRates, found: {repr(b12_val)}")
    except Exception as e:
        print(f"ERROR: Component 3 (tax rate lookup) — {e}")

    # Component 4: Tax amount formula in B13 (0.15 points)
    # Should be =B11*B12
    try:
        b13_val = ws['B13'].value
        b13_norm = normalize_formula(b13_val)
        if b13_norm in ("=B11*B12", "=B12*B11"):
            print(f"PASS: B13 has tax amount formula: {b13_val} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: B13 expected =B11*B12, found: {repr(b13_val)}")
    except Exception as e:
        print(f"ERROR: Component 4 (tax amount) — {e}")

    # Component 5: Grand total formula in B15 (0.15 points)
    # Should be =B11+B13+B14 (subtotal + tax + shipping)
    try:
        b15_val = ws['B15'].value
        b15_norm = normalize_formula(b15_val)
        # Accept any ordering of B11+B13+B14
        expected_terms = {"B11", "B13", "B14"}
        # Parse: remove '=' then split by '+'
        if b15_norm.startswith("="):
            parts = set(b15_norm[1:].split("+"))
            if parts == expected_terms:
                print(f"PASS: B15 has grand total formula: {b15_val} (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: B15 expected =B11+B13+B14, found: {repr(b15_val)}")
        else:
            print(f"FAIL: B15 is not a formula, found: {repr(b15_val)}")
    except Exception as e:
        print(f"ERROR: Component 5 (grand total) — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
