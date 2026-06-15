"""
Initial Setup: Invoice generator with line items, tax lookup, and grand total
Task ID: calc_sales_071
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_071'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Invoice ---
    ws = wb.active
    ws.title = 'Invoice'

    # Header info
    ws['A1'] = 'Invoice #: INV-2025-042'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = 'Customer: Acme Corp'
    ws['A2'].font = Font(size=12)
    ws['A3'] = 'State: CA'
    ws['A3'].font = Font(size=12)

    # Column headers row 5
    headers = {'A5': 'Item', 'B5': 'Qty', 'C5': 'Unit Price', 'D5': 'Line Total'}
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    for coord, val in headers.items():
        cell = ws[coord]
        cell.value = val
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align

    # Line items data (NO formulas - task asks agent to add them)
    items = [
        ('Widget Pro', 10, 150),
        ('Widget Basic', 25, 85),
        ('Support Plan', 1, 5000),
        ('Training', 3, 1200),
    ]
    for i, (item, qty, price) in enumerate(items):
        row = 6 + i
        ws.cell(row=row, column=1, value=item)
        ws.cell(row=row, column=2, value=qty)
        ws.cell(row=row, column=3, value=price)
        ws.cell(row=row, column=3).number_format = '$#,##0.00'
        # D column intentionally left empty - agent must add formulas

    # Summary labels
    ws['A11'] = 'Subtotal'
    ws['A11'].font = Font(bold=True)
    ws['A12'] = 'Tax Rate'
    ws['A12'].font = Font(bold=True)
    ws['A13'] = 'Tax Amount'
    ws['A13'].font = Font(bold=True)
    ws['A14'] = 'Shipping'
    ws['A14'].font = Font(bold=True)
    ws['B14'] = 250
    ws['B14'].number_format = '$#,##0.00'
    ws['A15'] = 'Grand Total'
    ws['A15'].font = Font(bold=True, size=12)

    # B11, B12, B13, B15 intentionally empty - agent must add formulas

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14

    # --- Sheet: TaxRates ---
    ws2 = wb.create_sheet('TaxRates')
    ws2['A1'] = 'State'
    ws2['B1'] = 'Rate'
    ws2['A1'].font = Font(bold=True)
    ws2['B1'].font = Font(bold=True)

    tax_data = [
        ('CA', 0.0725),
        ('TX', 0.0625),
        ('NY', 0.08),
    ]
    for i, (state, rate) in enumerate(tax_data):
        row = 2 + i
        ws2.cell(row=row, column=1, value=state)
        ws2.cell(row=row, column=2, value=rate)
        ws2.cell(row=row, column=2).number_format = '0.0000'

    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
