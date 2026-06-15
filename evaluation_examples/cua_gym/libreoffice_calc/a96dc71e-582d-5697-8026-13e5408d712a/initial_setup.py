"""
Initial Setup: Create a spreadsheet with grade distribution data for a class.
Task ID: calc_gpm_004
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_004'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Distribution ---
    ws = wb.active
    ws.title = 'Distribution'

    # Headers
    headers = ['Grade', 'Count']
    bold_font = Font(bold=True)
    bottom_border = Border(bottom=Side(style='thin', color='000000'))

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold_font
        cell.border = bottom_border

    # Data rows
    data = [
        ['A', 5],
        ['B', 8],
        ['C', 10],
        ['D', 4],
        ['F', 3],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
