"""
Reward Script: IFS/nested-IF performance tier classification in E2:E50
Task ID: calc_gg5_042
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30) - All 49 cells E2:E50 contain formulas
  Component 2 (0.35) - Formulas use IFS or nested IF with correct column refs (C, D)
  Component 3 (0.35) - Formula contains correct tier logic: Star (Score>=90 AND Tenure>=3),
                        Strong (>=80), Developing (>=65), At Risk (otherwise)
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_042'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Headcount' sheet must exist
    if 'Headcount' not in wb.sheetnames:
        print(f"CRITICAL: 'Headcount' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Headcount']

    # Component 1: All 49 cells E2:E50 contain formulas (0.30 points)
    try:
        formula_count = 0
        for r in range(2, 51):
            cell_val = ws.cell(row=r, column=5).value
            if isinstance(cell_val, str) and cell_val.startswith('='):
                formula_count += 1

        if formula_count == 49:
            print(f"PASS: Component 1 — All 49 cells E2:E50 contain formulas (0.30 pts)")
            total_score += 0.30
        elif formula_count > 0:
            # Partial credit proportional to coverage
            partial = round(0.30 * (formula_count / 49), 2)
            print(f"PARTIAL: Component 1 — {formula_count}/49 cells have formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No formula cells found in E2:E50 (0/49)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Formulas use IFS or nested IF and reference columns C and D (0.35 points)
    try:
        valid_structure_count = 0
        for r in range(2, 51):
            cell_val = ws.cell(row=r, column=5).value
            if not isinstance(cell_val, str) or not cell_val.startswith('='):
                continue
            upper_val = cell_val.upper().replace(' ', '')
            # Must use IFS(...) or nested IF(...IF(...))
            has_ifs = 'IFS(' in upper_val
            # Count nested IFs: at least 2 IF( occurrences for nested IF approach
            if_count = upper_val.count('IF(')
            has_nested_if = if_count >= 3  # nested IF with at least 3 levels
            # Must reference C column (Score) and D column (Tenure)
            has_c_ref = bool(re.search(r'C\d', cell_val, re.IGNORECASE))
            has_d_ref = bool(re.search(r'D\d', cell_val, re.IGNORECASE))
            if (has_ifs or has_nested_if) and has_c_ref and has_d_ref:
                valid_structure_count += 1

        if valid_structure_count == 49:
            print(f"PASS: Component 2 — All 49 formulas use IFS/nested-IF with C & D refs (0.35 pts)")
            total_score += 0.35
        elif valid_structure_count > 0:
            partial = round(0.35 * (valid_structure_count / 49), 2)
            print(f"PARTIAL: Component 2 — {valid_structure_count}/49 formulas have correct structure ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No formulas with correct IFS/nested-IF structure found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Formula contains correct tier logic (0.35 points)
    # Check for: Star (AND condition with >=90 and >=3), Strong (>=80), Developing (>=65), At Risk
    try:
        correct_logic_count = 0
        sample_formula = None
        for r in range(2, 51):
            cell_val = ws.cell(row=r, column=5).value
            if not isinstance(cell_val, str) or not cell_val.startswith('='):
                continue
            if sample_formula is None:
                sample_formula = cell_val

            upper_val = cell_val.upper().replace(' ', '')

            # Check for all four tier strings
            has_star = '"STAR"' in upper_val
            has_strong = '"STRONG"' in upper_val
            has_developing = '"DEVELOPING"' in upper_val
            has_at_risk = '"ATRISK"' in upper_val

            # Check for AND condition for Star (Score>=90 AND Tenure>=3)
            has_and_condition = 'AND(' in upper_val

            # Check for threshold values
            has_90 = '>=90' in upper_val or '>89' in upper_val
            has_80 = '>=80' in upper_val or '>79' in upper_val
            has_65 = '>=65' in upper_val or '>64' in upper_val
            has_3 = '>=3' in upper_val or '>2' in upper_val

            if (has_star and has_strong and has_developing and has_at_risk
                    and has_and_condition and has_90 and has_80 and has_65 and has_3):
                correct_logic_count += 1

        if correct_logic_count == 49:
            print(f"PASS: Component 3 — All 49 formulas have correct tier logic (0.35 pts)")
            print(f"  Sample formula: {sample_formula}")
            total_score += 0.35
        elif correct_logic_count > 0:
            partial = round(0.35 * (correct_logic_count / 49), 2)
            print(f"PARTIAL: Component 3 — {correct_logic_count}/49 formulas have correct logic ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No formulas with correct tier logic found")
            if sample_formula:
                print(f"  Sample formula found: {sample_formula}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook for LibreOffice
def persist_app_state(domain):
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
