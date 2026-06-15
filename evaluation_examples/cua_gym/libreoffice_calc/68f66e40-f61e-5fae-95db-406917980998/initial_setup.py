"""
Initial Setup: HR Analytics - Performance Tier Classification
Task ID: calc_gg5_042
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_042'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Headcount'

    # --- Headers ---
    headers = ['ID', 'Name', 'Score', 'Tenure', 'Tier']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border
        cell.font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")

    # --- Column widths ---
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 14

    # --- Employee Data (49 rows, rows 2-50) ---
    employees = [
        ('EMP001', 'Sarah Chen', 92, 5),
        ('EMP002', 'Marcus Johnson', 78, 2),
        ('EMP003', 'Priya Patel', 88, 4),
        ('EMP004', 'James O\'Brien', 61, 7),
        ('EMP005', 'Lucia Fernandez', 95, 1),
        ('EMP006', 'Wei Zhang', 73, 3),
        ('EMP007', 'Aisha Mohammed', 84, 6),
        ('EMP008', 'Robert Kim', 56, 2),
        ('EMP009', 'Elena Vasquez', 91, 4),
        ('EMP010', 'David Nakamura', 67, 8),
        ('EMP011', 'Fatima Al-Rashid', 82, 3),
        ('EMP012', 'Thomas Mueller', 45, 1),
        ('EMP013', 'Yuki Tanaka', 93, 6),
        ('EMP014', 'Olga Petrova', 76, 2),
        ('EMP015', 'Carlos Rivera', 69, 5),
        ('EMP016', 'Hannah Schmidt', 87, 3),
        ('EMP017', 'Raj Gupta', 52, 4),
        ('EMP018', 'Sophie Laurent', 90, 2),
        ('EMP019', 'Michael Brown', 81, 7),
        ('EMP020', 'Amara Okafor', 64, 1),
        ('EMP021', 'Kenji Watanabe', 96, 3),
        ('EMP022', 'Isabella Costa', 71, 5),
        ('EMP023', 'Ahmed Hassan', 83, 2),
        ('EMP024', 'Victoria Nguyen', 58, 6),
        ('EMP025', 'Daniel Park', 94, 4),
        ('EMP026', 'Marta Kowalski', 77, 3),
        ('EMP027', 'Samuel Osei', 66, 1),
        ('EMP028', 'Lena Johansson', 89, 8),
        ('EMP029', 'Christopher Lee', 43, 2),
        ('EMP030', 'Nadia Bouzid', 85, 5),
        ('EMP031', 'Ryan McCarthy', 72, 3),
        ('EMP032', 'Mei Lin Wang', 91, 7),
        ('EMP033', 'Fernando Silva', 60, 4),
        ('EMP034', 'Anna Bergman', 79, 2),
        ('EMP035', 'Tariq Mansour', 86, 6),
        ('EMP036', 'Julia Novak', 53, 1),
        ('EMP037', 'Hiroshi Yamamoto', 97, 5),
        ('EMP038', 'Grace Adebayo', 68, 3),
        ('EMP039', 'Patrick Sullivan', 80, 2),
        ('EMP040', 'Zara Khan', 74, 4),
        ('EMP041', 'Lucas Moreno', 92, 3),
        ('EMP042', 'Inna Volkov', 63, 7),
        ('EMP043', 'Benjamin Frost', 88, 1),
        ('EMP044', 'Chiara Romano', 55, 5),
        ('EMP045', 'Oscar Lindqvist', 90, 4),
        ('EMP046', 'Deepa Sharma', 81, 2),
        ('EMP047', 'William Chang', 70, 6),
        ('EMP048', 'Freya Andersen', 95, 3),
        ('EMP049', 'Ivan Petrov', 62, 1),
    ]

    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r, (emp_id, name, score, tenure) in enumerate(employees, 2):
        ws.cell(row=r, column=1, value=emp_id).border = data_border
        ws.cell(row=r, column=2, value=name).border = data_border
        c_score = ws.cell(row=r, column=3, value=score)
        c_score.border = data_border
        c_score.alignment = Alignment(horizontal="center")
        c_tenure = ws.cell(row=r, column=4, value=tenure)
        c_tenure.border = data_border
        c_tenure.alignment = Alignment(horizontal="center")
        # Column E (Tier) is intentionally left empty
        ws.cell(row=r, column=5).border = data_border

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
