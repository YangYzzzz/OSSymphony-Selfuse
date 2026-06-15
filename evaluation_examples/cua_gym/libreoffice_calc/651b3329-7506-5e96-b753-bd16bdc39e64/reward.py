"""
Reward Script: Food Nutrition and Meal Planning Tracker
Task ID: calc_grs_028
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): VLOOKUP formulas in Nutrition Analysis pulling data from Food Database
  Component 2 (0.15): Weekly total SUM formulas in column I
  Component 3 (0.20): Daily total formulas in rows 25-29
  Component 4 (0.20): Stacked bar chart with macronutrient breakdown (3 series)
  Component 5 (0.20): Conditional formatting on daily calorie totals (>2000 red, <1500 orange)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_028'


def persist_app_state(domain: str):
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: required sheets exist
    required_sheets = ['Food Database', 'Meal Plan', 'Nutrition Analysis']
    for sn in required_sheets:
        if sn not in wb.sheetnames:
            print(f"CRITICAL: Missing required sheet '{sn}'")
            print("REWARD: 0.0")
            return 0.0

    ws_na = wb['Nutrition Analysis']

    # =========================================================================
    # Component 1: VLOOKUP formulas in Nutrition Analysis B4:H23 (0.25 points)
    # These cells should contain VLOOKUP formulas referencing Food Database.
    # In the initial file, these cells are empty (no formulas).
    # =========================================================================
    try:
        vlookup_count = 0
        total_vlookup_cells = 0
        # Rows 4-23 (5 metrics x 4 meals = 20 rows), columns B-H (7 days)
        for row_num in range(4, 24):
            for col_num in range(2, 9):  # B=2 to H=8
                total_vlookup_cells += 1
                cell_val = ws_na.cell(row=row_num, column=col_num).value
                if cell_val is not None and isinstance(cell_val, str):
                    val_upper = cell_val.upper().replace(" ", "")
                    if "VLOOKUP" in val_upper or "INDEX" in val_upper or "MATCH" in val_upper:
                        vlookup_count += 1

        # Need at least 80% of cells to have lookup formulas
        vlookup_ratio = vlookup_count / total_vlookup_cells if total_vlookup_cells > 0 else 0
        if vlookup_ratio >= 0.8:
            print(f"PASS: Component 1 — VLOOKUP formulas found in {vlookup_count}/{total_vlookup_cells} cells ({vlookup_ratio:.0%}) (0.25 pts)")
            total_score += 0.25
        elif vlookup_ratio >= 0.4:
            partial = 0.25 * (vlookup_ratio / 0.8)
            print(f"PARTIAL: Component 1 — VLOOKUP formulas in {vlookup_count}/{total_vlookup_cells} cells ({vlookup_ratio:.0%}) ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Expected VLOOKUP/INDEX formulas in B4:H23, found {vlookup_count}/{total_vlookup_cells}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Weekly total SUM formulas in column I, rows 4-23 (0.15 points)
    # These cells should contain SUM formulas aggregating across the week.
    # In the initial file, column I rows 4-23 are empty.
    # =========================================================================
    try:
        sum_count = 0
        total_sum_cells = 0
        for row_num in range(4, 24):
            total_sum_cells += 1
            cell_val = ws_na.cell(row=row_num, column=9).value  # column I = 9
            if cell_val is not None and isinstance(cell_val, str):
                if "SUM" in cell_val.upper().replace(" ", ""):
                    sum_count += 1

        sum_ratio = sum_count / total_sum_cells if total_sum_cells > 0 else 0
        if sum_ratio >= 0.8:
            print(f"PASS: Component 2 — Weekly SUM formulas in {sum_count}/{total_sum_cells} column I cells (0.15 pts)")
            total_score += 0.15
        elif sum_ratio >= 0.3:
            partial = 0.15 * (sum_ratio / 0.8)
            print(f"PARTIAL: Component 2 — Weekly SUM formulas in {sum_count}/{total_sum_cells} cells ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Expected SUM formulas in I4:I23, found {sum_count}/{total_sum_cells}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Daily total formulas in rows 25-29, columns B-H (0.20 points)
    # Golden has formulas like =B4+B9+B14+B19 summing across meals for each day.
    # In the initial file, B25:H29 are empty.
    # =========================================================================
    try:
        daily_total_count = 0
        total_daily_cells = 0
        for row_num in range(25, 30):  # rows 25-29
            for col_num in range(2, 9):  # B-H
                total_daily_cells += 1
                cell_val = ws_na.cell(row=row_num, column=col_num).value
                if cell_val is not None and isinstance(cell_val, str):
                    # Should be a formula (starts with =)
                    if cell_val.startswith("="):
                        daily_total_count += 1

        daily_ratio = daily_total_count / total_daily_cells if total_daily_cells > 0 else 0
        if daily_ratio >= 0.8:
            print(f"PASS: Component 3 — Daily total formulas in {daily_total_count}/{total_daily_cells} cells (0.20 pts)")
            total_score += 0.20
        elif daily_ratio >= 0.3:
            partial = 0.20 * (daily_ratio / 0.8)
            print(f"PARTIAL: Component 3 — Daily total formulas in {daily_total_count}/{total_daily_cells} cells ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Expected formulas in B25:H29, found {daily_total_count}/{total_daily_cells}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: Stacked bar chart with macronutrient breakdown (0.20 points)
    # Golden has a BarChart with grouping='stacked' and 3 series (Protein, Carbs, Fat).
    # Initial file has 0 charts.
    # =========================================================================
    try:
        charts = ws_na._charts
        if len(charts) >= 1:
            chart = charts[0]
            is_bar = chart.__class__.__name__ == 'BarChart'
            is_stacked = getattr(chart, 'grouping', None) == 'stacked'
            has_series = len(chart.series) >= 3

            if is_bar and is_stacked and has_series:
                print(f"PASS: Component 4 — Stacked bar chart with {len(chart.series)} series (0.20 pts)")
                total_score += 0.20
            elif is_bar and has_series:
                # Bar chart with right number of series but not stacked
                print(f"PARTIAL: Component 4 — Bar chart with {len(chart.series)} series but grouping={getattr(chart, 'grouping', None)} (0.10 pts)")
                total_score += 0.10
            elif len(charts) >= 1:
                # Some chart exists but wrong type
                print(f"PARTIAL: Component 4 — Chart exists ({chart.__class__.__name__}) but expected stacked BarChart with 3+ series (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 4 — Chart found but not a valid stacked bar chart")
        else:
            print(f"FAIL: Component 4 — No charts found in Nutrition Analysis sheet")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # =========================================================================
    # Component 5: Conditional formatting on daily calorie totals (0.20 points)
    # Golden has CF on B25:H25: >2000 red fill, <1500 orange fill.
    # Initial file has 0 conditional formatting rules.
    # =========================================================================
    try:
        cf_rules = list(ws_na.conditional_formatting)
        if len(cf_rules) >= 1:
            found_red_high = False
            found_orange_low = False

            for cf in cf_rules:
                for rule in cf.rules:
                    rule_type = getattr(rule, 'type', '')
                    operator = getattr(rule, 'operator', '')
                    formula = getattr(rule, 'formula', [])

                    # Check for >2000 rule (red - calories exceed 2000)
                    if rule_type == 'cellIs' and operator == 'greaterThan':
                        if formula and '2000' in str(formula[0]):
                            found_red_high = True

                    # Check for <1500 rule (orange - calories below 1500)
                    if rule_type == 'cellIs' and operator == 'lessThan':
                        if formula and '1500' in str(formula[0]):
                            found_orange_low = True

            if found_red_high and found_orange_low:
                print(f"PASS: Component 5 — Conditional formatting with >2000 (red) and <1500 (orange) rules (0.20 pts)")
                total_score += 0.20
            elif found_red_high or found_orange_low:
                print(f"PARTIAL: Component 5 — Found {'red>2000' if found_red_high else ''} {'orange<1500' if found_orange_low else ''} (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 5 — CF rules exist but no matching >2000 or <1500 rules found")
        else:
            print(f"FAIL: Component 5 — No conditional formatting rules found in Nutrition Analysis sheet")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
