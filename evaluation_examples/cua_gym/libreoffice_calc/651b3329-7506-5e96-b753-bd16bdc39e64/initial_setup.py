"""
Initial Setup: Food Nutrition and Meal Planning Tracker
Task ID: calc_grs_028
Domain: libreoffice_calc

Creates a workbook with:
  Sheet1 "Food Database" - 20 food items with nutritional data and category dropdown
  Sheet2 "Meal Plan" - 7-day meal plan grid (Breakfast/Lunch/Dinner/Snack x Mon-Sun)
  Sheet3 "Nutrition Analysis" - Headers only, no formulas (task requires building them)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_028'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # === Shared styles ===
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # =====================================================
    # Sheet1: Food Database
    # =====================================================
    ws1 = wb.active
    ws1.title = "Food Database"

    headers1 = ["Food Name", "Serving Size", "Calories", "Protein (g)", "Carbs (g)", "Fat (g)", "Fiber (g)", "Category"]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 20 realistic food items
    foods = [
        ["Grilled Chicken Breast",  "150g",  248, 46.5, 0.0, 5.4, 0.0, "Protein"],
        ["Brown Rice",              "1 cup", 216, 5.0, 44.8, 1.8, 3.5, "Grain"],
        ["Steamed Broccoli",        "1 cup", 55,  3.7, 11.2, 0.6, 5.1, "Vegetable"],
        ["Salmon Fillet",           "170g",  367, 34.0, 0.0, 22.1, 0.0, "Protein"],
        ["Sweet Potato",            "1 medium", 103, 2.3, 24.0, 0.1, 3.8, "Vegetable"],
        ["Greek Yogurt",            "170g",  100, 17.0, 6.0, 0.7, 0.0, "Dairy"],
        ["Banana",                  "1 medium", 105, 1.3, 27.0, 0.4, 3.1, "Fruit"],
        ["Almonds",                 "30g",   164, 6.0, 6.1, 14.0, 3.5, "Fat"],
        ["Quinoa",                  "1 cup", 222, 8.1, 39.4, 3.6, 5.2, "Grain"],
        ["Spinach Salad",           "2 cups", 14,  1.7, 2.2, 0.2, 1.3, "Vegetable"],
        ["Whole Wheat Toast",       "2 slices", 138, 7.2, 24.0, 2.4, 3.8, "Grain"],
        ["Eggs (scrambled)",        "2 large", 182, 12.2, 2.4, 13.4, 0.0, "Protein"],
        ["Avocado",                 "1/2 medium", 120, 1.5, 6.4, 11.0, 5.0, "Fat"],
        ["Blueberries",             "1 cup", 84,  1.1, 21.4, 0.5, 3.6, "Fruit"],
        ["Cottage Cheese",          "1 cup", 206, 28.0, 6.2, 9.0, 0.0, "Dairy"],
        ["Turkey Sandwich",         "1 whole", 320, 24.0, 34.0, 8.0, 2.5, "Protein"],
        ["Mixed Vegetables",        "1 cup", 80,  4.0, 15.0, 0.5, 5.0, "Vegetable"],
        ["Apple",                   "1 medium", 95,  0.5, 25.1, 0.3, 4.4, "Fruit"],
        ["Oatmeal",                 "1 cup", 154, 5.3, 27.4, 2.6, 4.0, "Grain"],
        ["Peanut Butter",           "2 tbsp", 188, 7.0, 7.7, 16.0, 1.6, "Fat"],
    ]

    for r, row_data in enumerate(foods, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c in (3, 4, 5, 6, 7):  # numeric columns
                cell.number_format = '0.0'
                cell.alignment = Alignment(horizontal="center")
            elif c == 2:
                cell.alignment = Alignment(horizontal="center")

    # Category dropdown validation
    dv = DataValidation(
        type="list",
        formula1='"Protein,Vegetable,Fruit,Grain,Dairy,Fat"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "Please select a valid category"
    dv.errorTitle = "Invalid Category"
    dv.prompt = "Select food category"
    dv.promptTitle = "Category"
    dv.add("H2:H100")
    ws1.add_data_validation(dv)

    # Column widths
    ws1.column_dimensions["A"].width = 24
    ws1.column_dimensions["B"].width = 14
    ws1.column_dimensions["C"].width = 10
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["E"].width = 10
    ws1.column_dimensions["F"].width = 8
    ws1.column_dimensions["G"].width = 10
    ws1.column_dimensions["H"].width = 12

    ws1.freeze_panes = "A2"

    # =====================================================
    # Sheet2: Meal Plan
    # =====================================================
    ws2 = wb.create_sheet("Meal Plan")

    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    meals = ["Breakfast", "Lunch", "Dinner", "Snack"]

    # Header row: blank + days
    cell = ws2.cell(row=1, column=1, value="Meal")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

    for col, day in enumerate(days, 2):
        cell = ws2.cell(row=1, column=col, value=day)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Meal rows with food items from the database
    meal_plan = {
        "Breakfast": [
            "Oatmeal", "Eggs (scrambled)", "Greek Yogurt", "Whole Wheat Toast",
            "Oatmeal", "Eggs (scrambled)", "Greek Yogurt"
        ],
        "Lunch": [
            "Turkey Sandwich", "Grilled Chicken Breast", "Salmon Fillet", "Turkey Sandwich",
            "Quinoa", "Grilled Chicken Breast", "Turkey Sandwich"
        ],
        "Dinner": [
            "Salmon Fillet", "Brown Rice", "Grilled Chicken Breast", "Quinoa",
            "Salmon Fillet", "Sweet Potato", "Brown Rice"
        ],
        "Snack": [
            "Banana", "Almonds", "Apple", "Blueberries",
            "Peanut Butter", "Cottage Cheese", "Avocado"
        ],
    }

    for r, meal in enumerate(meals, 2):
        cell = ws2.cell(row=r, column=1, value=meal)
        cell.font = Font(name="Arial", size=11, bold=True)
        cell.fill = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border
        for col, food in enumerate(meal_plan[meal], 2):
            cell = ws2.cell(row=r, column=col, value=food)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws2.column_dimensions["A"].width = 12
    for col_letter in ["B", "C", "D", "E", "F", "G", "H"]:
        ws2.column_dimensions[col_letter].width = 22

    ws2.freeze_panes = "B2"

    # =====================================================
    # Sheet3: Nutrition Analysis (headers only, no formulas)
    # =====================================================
    ws3 = wb.create_sheet("Nutrition Analysis")

    # Title
    ws3.cell(row=1, column=1, value="Daily Nutrition Analysis")
    ws3.cell(row=1, column=1).font = Font(name="Arial", size=14, bold=True)

    # Column headers for day-by-day analysis
    analysis_headers = ["Metric", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday", "Weekly Total"]
    for col, h in enumerate(analysis_headers, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Row labels for each meal's nutrients
    nutrient_rows = [
        "Breakfast - Calories", "Breakfast - Protein (g)", "Breakfast - Carbs (g)", "Breakfast - Fat (g)", "Breakfast - Fiber (g)",
        "Lunch - Calories", "Lunch - Protein (g)", "Lunch - Carbs (g)", "Lunch - Fat (g)", "Lunch - Fiber (g)",
        "Dinner - Calories", "Dinner - Protein (g)", "Dinner - Carbs (g)", "Dinner - Fat (g)", "Dinner - Fiber (g)",
        "Snack - Calories", "Snack - Protein (g)", "Snack - Carbs (g)", "Snack - Fat (g)", "Snack - Fiber (g)",
        "",
        "Daily Total Calories", "Daily Total Protein (g)", "Daily Total Carbs (g)", "Daily Total Fat (g)", "Daily Total Fiber (g)",
    ]

    for r, label in enumerate(nutrient_rows, 4):
        cell = ws3.cell(row=r, column=1, value=label)
        cell.border = thin_border
        if label.startswith("Daily Total"):
            cell.font = Font(name="Arial", size=11, bold=True)
            cell.fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
        elif label == "":
            pass
        else:
            cell.font = Font(name="Arial", size=10)

    # Borders for the data area (empty, to be filled with formulas by the agent)
    for r in range(4, 4 + len(nutrient_rows)):
        for c in range(2, 10):
            ws3.cell(row=r, column=c).border = thin_border

    ws3.column_dimensions["A"].width = 28
    for col_letter in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        ws3.column_dimensions[col_letter].width = 14

    # Note at bottom telling user what to do
    note_row = 4 + len(nutrient_rows) + 2
    ws3.cell(row=note_row, column=1, value="Instructions: Use VLOOKUP to pull nutritional data from the Food Database for each meal entry in the Meal Plan.")
    ws3.cell(row=note_row, column=1).font = Font(name="Arial", size=10, italic=True, color="808080")

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
