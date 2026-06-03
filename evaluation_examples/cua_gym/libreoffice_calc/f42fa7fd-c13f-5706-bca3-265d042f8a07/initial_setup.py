"""
Initial Setup: Workout log and fitness tracker
Task ID: calc_wf_033
Domain: libreoffice_calc

Creates a workout log with 8 weeks of exercise data (no formulas, no charts,
no conditional formatting). Volume column is left empty for the agent to fill.
PRs sheet has exercise names but no formulas.
"""

import os
import shlex
import subprocess
import time
import random
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_033'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 'Log' ---
    ws_log = wb.active
    ws_log.title = 'Log'

    # Headers
    headers = ['Date', 'Exercise', 'Sets', 'Reps', 'Weight', 'Volume']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_font_white = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws_log.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Exercise pool (realistic compound + isolation movements)
    exercises = [
        'Barbell Squat', 'Bench Press', 'Deadlift', 'Overhead Press', 'Barbell Row',
        'Pull-Up', 'Dumbbell Curl', 'Tricep Dip', 'Leg Press', 'Romanian Deadlift',
        'Incline Bench Press', 'Lat Pulldown', 'Seated Row', 'Leg Curl', 'Calf Raise',
    ]

    # Session templates: each session uses 5 exercises
    session_templates = [
        ['Barbell Squat', 'Bench Press', 'Barbell Row', 'Dumbbell Curl', 'Tricep Dip'],          # Push/Pull A
        ['Deadlift', 'Overhead Press', 'Lat Pulldown', 'Leg Curl', 'Calf Raise'],                # Push/Pull B
        ['Leg Press', 'Incline Bench Press', 'Seated Row', 'Dumbbell Curl', 'Romanian Deadlift'],  # Upper/Lower A
        ['Barbell Squat', 'Bench Press', 'Pull-Up', 'Overhead Press', 'Barbell Row'],              # Compound Focus
    ]

    # Weight ranges per exercise (lbs)
    weight_ranges = {
        'Barbell Squat': (185, 275),
        'Bench Press': (135, 225),
        'Deadlift': (225, 365),
        'Overhead Press': (95, 155),
        'Barbell Row': (135, 205),
        'Pull-Up': (0, 45),      # bodyweight + added weight
        'Dumbbell Curl': (25, 50),
        'Tricep Dip': (0, 45),
        'Leg Press': (270, 450),
        'Romanian Deadlift': (135, 245),
        'Incline Bench Press': (115, 195),
        'Lat Pulldown': (120, 180),
        'Seated Row': (120, 180),
        'Leg Curl': (80, 140),
        'Calf Raise': (135, 225),
    }

    random.seed(42)  # reproducible data
    start_date = datetime(2025, 11, 3)  # Monday of week 1

    row = 2
    # Session days within a week: Mon, Wed, Fri, Sat
    session_days = [0, 2, 4, 5]

    for week in range(8):
        week_start = start_date + timedelta(weeks=week)
        for s_idx, day_offset in enumerate(session_days):
            session_date = week_start + timedelta(days=day_offset)
            template = session_templates[s_idx % len(session_templates)]

            for exercise in template:
                lo, hi = weight_ranges[exercise]
                # Progressive overload: slight increase over weeks
                base_weight = lo + (hi - lo) * (week / 10.0)
                weight = round(base_weight + random.uniform(-10, 10))
                weight = max(lo, min(hi, weight))
                # Round weight to nearest 5
                weight = round(weight / 5) * 5

                sets = random.choice([3, 4, 5])
                reps = random.choice([5, 6, 8, 10, 12])

                ws_log.cell(row=row, column=1, value=session_date.strftime('%Y-%m-%d'))
                ws_log.cell(row=row, column=2, value=exercise)
                ws_log.cell(row=row, column=3, value=sets)
                ws_log.cell(row=row, column=4, value=reps)
                ws_log.cell(row=row, column=5, value=weight)
                # Column F (Volume) intentionally left EMPTY
                row += 1

    last_data_row = row - 1

    # Set column widths
    ws_log.column_dimensions['A'].width = 14
    ws_log.column_dimensions['B'].width = 22
    ws_log.column_dimensions['C'].width = 8
    ws_log.column_dimensions['D'].width = 8
    ws_log.column_dimensions['E'].width = 10
    ws_log.column_dimensions['F'].width = 12

    # --- Sheet 'PRs' ---
    ws_prs = wb.create_sheet('PRs')

    pr_headers = ['Exercise', 'Best Weight', 'Best Volume', 'Date Achieved']
    for col, h in enumerate(pr_headers, 1):
        cell = ws_prs.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # List all unique exercises
    unique_exercises = [
        'Barbell Squat', 'Bench Press', 'Deadlift', 'Overhead Press', 'Barbell Row',
        'Pull-Up', 'Dumbbell Curl', 'Tricep Dip', 'Leg Press', 'Romanian Deadlift',
        'Incline Bench Press', 'Lat Pulldown', 'Seated Row', 'Leg Curl', 'Calf Raise',
    ]

    for i, ex in enumerate(unique_exercises, 2):
        ws_prs.cell(row=i, column=1, value=ex)
        # Columns B, C, D intentionally left EMPTY for MAXIFS formulas

    ws_prs.column_dimensions['A'].width = 22
    ws_prs.column_dimensions['B'].width = 14
    ws_prs.column_dimensions['C'].width = 14
    ws_prs.column_dimensions['D'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Log sheet: {last_data_row} data rows (row 2 to {last_data_row})')
    print(f'PRs sheet: {len(unique_exercises)} exercises listed')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
