"""
Reward Script: Workout Log and Fitness Tracker
Task ID: calc_wf_033
Domain: libreoffice_calc
Scoring:
  Component 1 — Volume formulas in Log!F column (0.25 pts)
  Component 2 — MAXIFS formulas in PRs sheet B/C columns (0.25 pts)
  Component 3 — Weekly volume summary section with aggregation formulas (0.15 pts)
  Component 4 — Line chart on Log sheet for weekly volume (0.15 pts)
  Component 5 — Conditional formatting on E column highlighting PRs in gold (0.20 pts)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_033'


def persist_app_state(domain: str):
    """Best-effort save via Ctrl+S for GUI apps."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify required sheets exist
    if 'Log' not in wb.sheetnames:
        print("CRITICAL: 'Log' sheet not found")
        print("REWARD: 0.0")
        return 0.0
    if 'PRs' not in wb.sheetnames:
        print("CRITICAL: 'PRs' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws_log = wb['Log']
    ws_prs = wb['PRs']

    # =========================================================================
    # Component 1: Volume formulas in Log!F column (0.25 points)
    # The task requires Volume = Sets * Reps * Weight in column F.
    # Initial state: F column is empty (None). Golden: formulas like =C*D*E.
    # =========================================================================
    try:
        volume_formula_count = 0
        total_data_rows = 0
        # Check a representative sample of rows (2 through at least 161)
        for r in range(2, ws_log.max_row + 1):
            # Skip rows that are part of weekly summary (after data ends)
            if ws_log.cell(row=r, column=2).value is None and ws_log.cell(row=r, column=3).value is None:
                continue
            # Only count actual exercise data rows (have exercise name in B)
            b_val = ws_log.cell(row=r, column=2).value
            if b_val is None:
                continue
            total_data_rows += 1
            f_val = ws_log.cell(row=r, column=6).value
            if f_val is not None and isinstance(f_val, str):
                # Accept any multiplication formula involving C, D, E columns
                f_upper = f_val.upper().replace(" ", "")
                if re.search(r'=.*C\d+.*\*.*D\d+.*\*.*E\d+', f_upper) or \
                   re.search(r'=.*D\d+.*\*.*C\d+.*\*.*E\d+', f_upper) or \
                   re.search(r'=.*E\d+.*\*.*C\d+.*\*.*D\d+', f_upper) or \
                   'PRODUCT' in f_upper:
                    volume_formula_count += 1

        if total_data_rows > 0:
            coverage = volume_formula_count / total_data_rows
            if coverage >= 0.9:
                print(f"PASS: Component 1 — Volume formulas in {volume_formula_count}/{total_data_rows} data rows ({coverage:.0%}) (0.25 pts)")
                total_score += 0.25
            elif coverage >= 0.5:
                partial = 0.25 * coverage
                print(f"PARTIAL: Component 1 — Volume formulas in {volume_formula_count}/{total_data_rows} rows ({coverage:.0%}) ({partial:.3f} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 1 — Only {volume_formula_count}/{total_data_rows} rows have volume formulas ({coverage:.0%})")
        else:
            print("FAIL: Component 1 — No data rows found in Log sheet")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: MAXIFS formulas in PRs sheet B and C columns (0.25 points)
    # Initial state: PRs B/C/D columns are empty. Golden: MAXIFS in B/C.
    # =========================================================================
    try:
        maxifs_count = 0
        pr_rows = 0
        for r in range(2, ws_prs.max_row + 1):
            exercise = ws_prs.cell(row=r, column=1).value
            if exercise is None:
                continue
            pr_rows += 1
            b_val = ws_prs.cell(row=r, column=2).value
            c_val = ws_prs.cell(row=r, column=3).value
            b_ok = False
            c_ok = False
            if b_val is not None and isinstance(b_val, str):
                b_upper = b_val.upper().replace(" ", "")
                if 'MAXIFS' in b_upper or 'MAX(' in b_upper or 'MAXIF' in b_upper:
                    b_ok = True
            if c_val is not None and isinstance(c_val, str):
                c_upper = c_val.upper().replace(" ", "")
                if 'MAXIFS' in c_upper or 'MAX(' in c_upper or 'MAXIF' in c_upper:
                    c_ok = True
            if b_ok and c_ok:
                maxifs_count += 1

        if pr_rows > 0:
            coverage = maxifs_count / pr_rows
            if coverage >= 0.9:
                print(f"PASS: Component 2 — MAXIFS formulas in {maxifs_count}/{pr_rows} PR rows for both Best Weight and Best Volume (0.25 pts)")
                total_score += 0.25
            elif coverage >= 0.5:
                partial = 0.25 * coverage
                print(f"PARTIAL: Component 2 — MAXIFS in {maxifs_count}/{pr_rows} PR rows ({partial:.3f} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 2 — Only {maxifs_count}/{pr_rows} PR rows have MAXIFS formulas")
        else:
            print("FAIL: Component 2 — No exercise rows in PRs sheet")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Weekly volume summary section (0.15 points)
    # Golden has rows 163-172 with weekly summary using SUMPRODUCT/SUMIFS/WEEKNUM.
    # Initial has nothing beyond row 161.
    # =========================================================================
    try:
        weekly_summary_found = False
        weekly_formula_count = 0
        # Search for weekly summary section anywhere after the main data
        for r in range(2, ws_log.max_row + 1):
            val = ws_log.cell(row=r, column=1).value
            if val is not None and isinstance(val, str) and 'WEEKLY' in val.upper() and 'VOLUME' in val.upper():
                weekly_summary_found = True
            # Check for weekly aggregation formulas (SUMPRODUCT, SUMIFS, etc.)
            for c in range(1, ws_log.max_column + 1):
                cell_val = ws_log.cell(row=r, column=c).value
                if cell_val is not None and isinstance(cell_val, str):
                    cv_upper = cell_val.upper()
                    if ('SUMPRODUCT' in cv_upper or 'SUMIFS' in cv_upper) and \
                       ('WEEKNUM' in cv_upper or 'WEEK' in cv_upper):
                        weekly_formula_count += 1

        if weekly_formula_count >= 8:
            print(f"PASS: Component 3 — Found {weekly_formula_count} weekly aggregation formulas (0.15 pts)")
            total_score += 0.15
        elif weekly_formula_count >= 4:
            partial = 0.15 * (weekly_formula_count / 8)
            print(f"PARTIAL: Component 3 — Found {weekly_formula_count}/8 weekly formulas ({partial:.3f} pts)")
            total_score += partial
        elif weekly_summary_found:
            print(f"PARTIAL: Component 3 — Weekly summary header found but only {weekly_formula_count} aggregation formulas (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 3 — No weekly volume summary section found (formulas: {weekly_formula_count})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: Line chart on Log sheet (0.15 points)
    # Initial: 0 charts. Golden: 1 line chart for weekly total volume.
    # =========================================================================
    try:
        charts = ws_log._charts
        if len(charts) >= 1:
            # Check if any chart is a line chart
            line_chart_found = False
            for chart in charts:
                chart_class = chart.__class__.__name__
                if 'Line' in chart_class:
                    line_chart_found = True
                    break
            if line_chart_found:
                print(f"PASS: Component 4 — Line chart found on Log sheet ({len(charts)} chart(s)) (0.15 pts)")
                total_score += 0.15
            else:
                # There is a chart but not a line chart — partial credit
                print(f"PARTIAL: Component 4 — Chart found but not a line chart (class: {charts[0].__class__.__name__}) (0.08 pts)")
                total_score += 0.08
        else:
            # Also check other sheets for charts
            any_chart = False
            for sname in wb.sheetnames:
                if len(wb[sname]._charts) > 0:
                    any_chart = True
                    break
            if any_chart:
                print(f"PARTIAL: Component 4 — Chart found on another sheet, not Log (0.05 pts)")
                total_score += 0.05
            else:
                print("FAIL: Component 4 — No charts found in workbook")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # =========================================================================
    # Component 5: Conditional formatting on weight column highlighting PRs (0.20 points)
    # Initial: no conditional formatting. Golden: expression rule on E2:E161
    # with formula checking if weight = MAXIFS (best weight for that exercise),
    # highlighted in gold (FFFFD700).
    # =========================================================================
    try:
        cf_rules = list(ws_log.conditional_formatting)
        if len(cf_rules) > 0:
            pr_highlight_found = False
            for cf in cf_rules:
                cf_range = str(cf)
                for rule in cf.rules:
                    # Check if it references MAXIFS or MAX in the formula
                    formulas = rule.formula if hasattr(rule, 'formula') else []
                    formula_text = ' '.join(str(f) for f in formulas).upper() if formulas else ''
                    # The golden uses: E2=MAXIFS(E$2:E$161,B$2:B$161,B2)
                    if ('MAXIFS' in formula_text or 'MAX' in formula_text) and \
                       ('E' in cf_range.upper()):
                        pr_highlight_found = True
                        # Bonus: check for gold color
                        has_gold = False
                        if hasattr(rule, 'dxf') and rule.dxf:
                            if rule.dxf.fill and rule.dxf.fill.fgColor:
                                color = rule.dxf.fill.fgColor.rgb
                                if color and 'FFD700' in str(color).upper():
                                    has_gold = True
                        if has_gold:
                            print(f"PASS: Component 5 — Conditional formatting with MAXIFS and gold fill on {cf_range} (0.20 pts)")
                            total_score += 0.20
                        else:
                            print(f"PARTIAL: Component 5 — Conditional formatting with MAXIFS on {cf_range} but no gold fill (0.15 pts)")
                            total_score += 0.15
                        break
                if pr_highlight_found:
                    break

            if not pr_highlight_found:
                # Some conditional formatting exists but doesn't match PR pattern
                print(f"PARTIAL: Component 5 — Conditional formatting found but does not use MAXIFS for PR highlighting (0.05 pts)")
                total_score += 0.05
        else:
            print("FAIL: Component 5 — No conditional formatting rules found on Log sheet")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
