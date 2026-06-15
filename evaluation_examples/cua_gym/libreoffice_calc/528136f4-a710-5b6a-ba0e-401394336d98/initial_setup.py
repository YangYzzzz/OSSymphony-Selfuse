"""
Initial Setup: Configure a stacked bar chart on the 'Budget Overview' sheet
Task ID: calc_gg2_033
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg2_033'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Budget Overview ---
    ws = wb.active
    ws.title = "Budget Overview"

    # Title row
    ws.merge_cells("A1:D1")
    ws["A1"] = "FY2025 Budget Allocation Report"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Subtitle
    ws.merge_cells("A2:D2")
    ws["A2"] = "Prepared by CFO Office - Q1 Review"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Column headers in row 4
    headers = ["Department", "Personnel", "Infrastructure", "Programs"]
    header_fill = PatternFill(start_color="FF1F4E79", end_color="FF1F4E79", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows (in thousands $)
    data = [
        ["HR", 285000, 42000, 118000],
        ["IT", 410000, 195000, 87000],
        ["Marketing", 195000, 68000, 312000],
        ["Operations", 520000, 248000, 145000],
        ["Finance", 310000, 56000, 92000],
    ]

    data_font = Font(name="Arial", size=11)
    money_format = '$#,##0'
    for r, row_data in enumerate(data, 5):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = thin_border
            if c == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = Font(name="Arial", size=11, bold=True)
            else:
                cell.number_format = money_format
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # Total row
    ws.cell(row=10, column=1, value="Total").font = Font(name="Arial", size=11, bold=True, italic=True)
    ws.cell(row=10, column=1).border = thin_border
    for col_idx in range(2, 5):
        cell = ws.cell(row=10, column=col_idx)
        col_letter = chr(64 + col_idx)
        cell.value = f"=SUM({col_letter}5:{col_letter}9)"
        cell.number_format = money_format
        cell.font = Font(name="Arial", size=11, bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="right", vertical="center")

    # Set column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14

    # --- Create Clustered Column Chart (the INITIAL state) ---
    chart = BarChart()
    chart.type = "col"  # vertical columns
    chart.grouping = "clustered"
    chart.title = "Budget Allocation by Department"
    chart.y_axis.title = "Amount ($)"
    chart.x_axis.title = "Department"
    chart.style = 10
    chart.width = 20
    chart.height = 14

    # Data: columns B-D (Personnel, Infrastructure, Programs), rows 4-9 (header + 5 depts)
    data_ref = Reference(ws, min_col=2, min_row=4, max_col=4, max_row=9)
    cats_ref = Reference(ws, min_col=1, min_row=5, max_row=9)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    # Set series colors
    from openpyxl.chart.series import DataPoint
    from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
    colors = ["4472C4", "ED7D31", "70AD47"]
    for i, color in enumerate(colors):
        chart.series[i].graphicalProperties.solidFill = color

    ws.add_chart(chart, "A12")

    # --- Additional sheet: Raw Data Notes ---
    ws2 = wb.create_sheet("Data Notes")
    ws2["A1"] = "Budget Data Notes"
    ws2["A1"].font = Font(name="Arial", size=12, bold=True)
    notes = [
        ["Note", "Description", "Date"],
        ["1", "Personnel costs include salaries, benefits, and training", "2025-01-10"],
        ["2", "Infrastructure covers office space, equipment, and IT systems", "2025-01-10"],
        ["3", "Programs include all departmental projects and initiatives", "2025-01-15"],
        ["4", "All figures are FY2025 approved budget amounts", "2025-02-01"],
        ["5", "Q2 adjustments pending board approval", "2025-03-20"],
    ]
    for r, row_data in enumerate(notes, 3):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if r == 3:
                cell.font = Font(bold=True)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
