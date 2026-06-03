"""
Initial Setup: Inventory reorder calculator - raw data only
Task ID: calc_wf_087
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_087'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Inventory'

    # Headers
    headers = [
        'SKU', 'Product Name', 'Current Stock', 'Daily Usage',
        'Lead Time (days)', 'Safety Stock', 'Unit Cost ($)',
        'Order Cost ($)', 'Holding Cost (%/year)'
    ]
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 25 products with realistic inventory data
    # Columns: SKU, Name, Current Stock, Daily Usage, Lead Time, Safety Stock, Unit Cost, Order Cost, Holding Cost%
    products = [
        ['WH-1001', 'Stainless Steel Bolts M8', 4500, 120, 7, 200, 0.35, 45.00, 0.20],
        ['WH-1002', 'Copper Wire Spool 14AWG', 180, 8, 14, 30, 28.50, 120.00, 0.18],
        ['WH-1003', 'PVC Pipe 2in x 10ft', 320, 15, 10, 50, 12.75, 85.00, 0.15],
        ['WH-1004', 'LED Panel Light 40W', 95, 6, 21, 40, 42.00, 150.00, 0.22],
        ['WH-1005', 'Hydraulic Hose 1/2in', 210, 12, 12, 45, 18.90, 95.00, 0.20],
        ['WH-1006', 'Aluminum Sheet 4x8ft', 45, 3, 18, 15, 185.00, 250.00, 0.16],
        ['WH-1007', 'Rubber Gasket Set', 1200, 45, 8, 120, 2.50, 60.00, 0.25],
        ['WH-1008', 'Industrial Adhesive 500ml', 380, 20, 5, 35, 15.60, 55.00, 0.18],
        ['WH-1009', 'Welding Rod E6013 5kg', 150, 7, 15, 30, 22.00, 110.00, 0.20],
        ['WH-1010', 'Circuit Breaker 20A', 85, 4, 20, 25, 35.00, 130.00, 0.22],
        ['WH-1011', 'Nylon Cable Ties 300mm', 8500, 300, 6, 600, 0.08, 35.00, 0.15],
        ['WH-1012', 'Bearing 6205-2RS', 240, 10, 14, 40, 8.75, 75.00, 0.20],
        ['WH-1013', 'Spray Paint Matte Black', 420, 18, 7, 45, 6.50, 50.00, 0.18],
        ['WH-1014', 'Stainless Pipe Fitting 1in', 160, 9, 16, 50, 14.25, 90.00, 0.20],
        ['WH-1015', 'Thermal Insulation Roll', 75, 2, 25, 15, 95.00, 200.00, 0.14],
        ['WH-1016', 'Safety Gloves Nitrile L', 3200, 150, 5, 250, 0.45, 40.00, 0.22],
        ['WH-1017', 'Pressure Gauge 0-100PSI', 55, 3, 18, 18, 28.00, 105.00, 0.20],
        ['WH-1018', 'Silicone Sealant 300ml', 600, 25, 6, 55, 7.80, 48.00, 0.18],
        ['WH-1019', 'Drill Bit Set HSS 1-13mm', 90, 5, 12, 20, 32.00, 85.00, 0.22],
        ['WH-1020', 'Conveyor Belt Segment 1m', 30, 1, 30, 10, 245.00, 350.00, 0.12],
        ['WH-1021', 'Fluorescent Tube T8 4ft', 500, 22, 8, 60, 4.20, 55.00, 0.16],
        ['WH-1022', 'Steel Chain 3/8in per ft', 1800, 65, 10, 200, 1.85, 70.00, 0.18],
        ['WH-1023', 'Air Filter Element 10in', 130, 6, 14, 28, 16.50, 80.00, 0.20],
        ['WH-1024', 'Epoxy Resin Kit 1L', 200, 8, 9, 25, 24.00, 95.00, 0.22],
        ['WH-1025', 'Vibration Dampener Pad', 70, 2, 22, 15, 55.00, 140.00, 0.16],
    ]

    for r, row_data in enumerate(products, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 1:  # SKU
                cell.alignment = Alignment(horizontal='center')
            elif c >= 3:  # numeric columns
                cell.alignment = Alignment(horizontal='right')
                if c == 7:  # Unit Cost
                    cell.number_format = '$#,##0.00'
                elif c == 8:  # Order Cost
                    cell.number_format = '$#,##0.00'
                elif c == 9:  # Holding Cost %
                    cell.number_format = '0%'

    # Set column widths
    col_widths = [12, 28, 14, 12, 16, 13, 14, 14, 18]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[chr(65 + i)].width = w

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
