"""
Reward Script: Inventory Reorder Calculator
Task ID: calc_wf_087
Domain: libreoffice_calc
Scoring:
  Component 1: Reorder Point formulas in column J (0.20 pts)
  Component 2: EOQ formulas in column K (0.20 pts)
  Component 3: Days of Supply formulas in column L (0.15 pts)
  Component 4: Order Urgency formulas in column M (0.15 pts)
  Component 5: Conditional formatting on Order Urgency column (0.15 pts)
  Component 6: Bar chart for Days of Supply (0.15 pts)
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_087'
FILE_PATH = os.path.join(WORKDIR, f'{TASK_ID}.xlsx')

DATA_ROWS = range(2, 27)  # rows 2-26 (25 products)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that 'Inventory' sheet exists
    if 'Inventory' not in wb.sheetnames:
        print("CRITICAL: 'Inventory' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Inventory']

    # Component 1: Reorder Point formulas in column J (0.20 points)
    # Formula pattern: =(D*E)+F  i.e. (Daily Usage * Lead Time) + Safety Stock
    try:
        rp_count = 0
        for r in DATA_ROWS:
            val = ws.cell(row=r, column=10).value  # Column J
            if val is not None and isinstance(val, str):
                v = val.upper().replace(" ", "")
                # Check for pattern like =(Dx*Ex)+Fx or =Dx*Ex+Fx
                if ('D' in v and 'E' in v and 'F' in v and
                        ('*' in v) and ('+' in v or '&' in v)):
                    rp_count += 1
                # Also accept SUMPRODUCT or other valid approaches
                elif 'SUMPRODUCT' in v or 'SUM' in v:
                    rp_count += 1
        # Award points proportionally: need at least 20 of 25 rows
        if rp_count >= 20:
            print(f"PASS: Component 1 - Reorder Point formulas found in {rp_count}/25 rows (0.20 pts)")
            total_score += 0.20
        elif rp_count >= 10:
            partial = 0.10
            print(f"PARTIAL: Component 1 - Reorder Point formulas in {rp_count}/25 rows ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - Reorder Point formulas found in only {rp_count}/25 rows")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: EOQ formulas in column K (0.20 points)
    # Formula pattern: =SQRT(2*(D*365)*H/(G*I))
    try:
        eoq_count = 0
        for r in DATA_ROWS:
            val = ws.cell(row=r, column=11).value  # Column K
            if val is not None and isinstance(val, str):
                v = val.upper().replace(" ", "")
                if 'SQRT' in v:
                    eoq_count += 1
        if eoq_count >= 20:
            print(f"PASS: Component 2 - EOQ formulas (SQRT) found in {eoq_count}/25 rows (0.20 pts)")
            total_score += 0.20
        elif eoq_count >= 10:
            partial = 0.10
            print(f"PARTIAL: Component 2 - EOQ formulas in {eoq_count}/25 rows ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 - EOQ formulas found in only {eoq_count}/25 rows")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Days of Supply formulas in column L (0.15 points)
    # Formula pattern: =C/D (Current Stock / Daily Usage)
    try:
        dos_count = 0
        for r in DATA_ROWS:
            val = ws.cell(row=r, column=12).value  # Column L
            if val is not None and isinstance(val, str):
                v = val.upper().replace(" ", "")
                # Pattern: =Cx/Dx  (current stock / daily usage)
                if '/' in v and 'C' in v and 'D' in v:
                    dos_count += 1
        if dos_count >= 20:
            print(f"PASS: Component 3 - Days of Supply formulas found in {dos_count}/25 rows (0.15 pts)")
            total_score += 0.15
        elif dos_count >= 10:
            partial = 0.07
            print(f"PARTIAL: Component 3 - Days of Supply formulas in {dos_count}/25 rows ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 - Days of Supply formulas found in only {dos_count}/25 rows")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Order Urgency formulas in column M (0.15 points)
    # Formula pattern: =IF(C<=J,"ORDER NOW",IF(C<=J*1.2,"SOON","OK"))
    try:
        urg_count = 0
        for r in DATA_ROWS:
            val = ws.cell(row=r, column=13).value  # Column M
            if val is not None and isinstance(val, str):
                v = val.upper().replace(" ", "")
                # Must be an IF formula containing ORDER NOW or ORDERNOW
                if v.startswith('=') and 'IF' in v and ('ORDERNOW' in v.replace(' ', '') or 'ORDER' in v):
                    urg_count += 1
        if urg_count >= 20:
            print(f"PASS: Component 4 - Order Urgency formulas found in {urg_count}/25 rows (0.15 pts)")
            total_score += 0.15
        elif urg_count >= 10:
            partial = 0.07
            print(f"PARTIAL: Component 4 - Order Urgency formulas in {urg_count}/25 rows ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 - Order Urgency formulas found in only {urg_count}/25 rows")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Conditional formatting on Order Urgency column M (0.15 points)
    # Expect rules on M2:M26 area — red for ORDER NOW, yellow for SOON
    try:
        cf_list = list(ws.conditional_formatting)
        order_now_count = 0
        soon_count = 0
        for cfr in cf_list:
            range_str = str(cfr)
            # Check if the range covers column M (at least partially)
            if 'M' in range_str or 'm' in range_str:
                for rule in cfr.rules:
                    formula_str = str(getattr(rule, 'formula', []))
                    if 'ORDER NOW' in formula_str or 'ORDER' in formula_str:
                        order_now_count += 1
                    if 'SOON' in formula_str:
                        soon_count += 1

        if order_now_count > 0 and soon_count > 0:
            print(f"PASS: Component 5 - Conditional formatting with ORDER NOW and SOON rules (0.15 pts)")
            total_score += 0.15
        elif order_now_count > 0 or soon_count > 0:
            partial = 0.07
            found = "ORDER NOW" if order_now_count > 0 else "SOON"
            print(f"PARTIAL: Component 5 - Only {found} conditional formatting rule found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 - No conditional formatting rules found on column M")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    # Component 6: Bar chart for Days of Supply (0.15 points)
    # Expect at least one BarChart in the Inventory sheet
    try:
        charts = ws._charts
        bar_chart_count = sum(1 for ch in charts if ch.__class__.__name__ == 'BarChart')

        if bar_chart_count > 0:
            print(f"PASS: Component 6 - Bar chart found in Inventory sheet (0.15 pts)")
            total_score += 0.15
        else:
            # Check all sheets for a bar chart
            for sn in wb.sheetnames:
                bar_chart_count += sum(1 for ch in wb[sn]._charts if ch.__class__.__name__ == 'BarChart')
            if bar_chart_count > 0:
                print(f"PASS: Component 6 - Bar chart found (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 6 - No bar chart found in any sheet")
    except Exception as e:
        print(f"ERROR: Component 6 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
if not os.path.exists(FILE_PATH):
    print(f"File not found: {FILE_PATH}")
    print("REWARD: 0.0")
else:
    verify_task(FILE_PATH)
