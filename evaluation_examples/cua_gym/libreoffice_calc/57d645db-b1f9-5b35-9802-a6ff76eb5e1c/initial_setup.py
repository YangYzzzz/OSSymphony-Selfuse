"""
Initial Setup: Marketing campaign performance spreadsheet with raw data and formulas.
Task ID: calc_gsd_048
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_048'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Campaigns'

    # --- Headers in row 1 (NOT bold - task asks agent to bold them) ---
    headers = [
        'Campaign ID', 'Campaign Name', 'Spend', 'Impressions', 'Clicks',
        'Conversions', 'Revenue', 'CTR', 'Conversion Rate', 'ROAS'
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # --- 20 realistic campaign rows (rows 2-21) ---
    campaigns = [
        ['MKT-001', 'Spring Brand Awareness', 12500, 850000, 17000, 510, 38250],
        ['MKT-002', 'Summer Sale Push', 8700, 620000, 14880, 595, 41650],
        ['MKT-003', 'Product Launch Alpha', 25000, 1200000, 30000, 900, 112500],
        ['MKT-004', 'Holiday Retargeting', 15300, 980000, 22540, 676, 50700],
        ['MKT-005', 'Q1 Email Drip', 3200, 210000, 6300, 252, 15120],
        ['MKT-006', 'Social Media Blitz', 9800, 730000, 16790, 336, 20160],
        ['MKT-007', 'Influencer Partnership', 18500, 1450000, 37700, 754, 67860],
        ['MKT-008', 'Back to School Promo', 7400, 490000, 11270, 338, 21972],
        ['MKT-009', 'Black Friday Deals', 32000, 2100000, 63000, 1890, 170100],
        ['MKT-010', 'Year-End Clearance', 14200, 920000, 20240, 607, 36420],
        ['MKT-011', 'New Customer Acquisition', 21000, 1580000, 39500, 790, 55300],
        ['MKT-012', 'Loyalty Program Boost', 5600, 380000, 9880, 395, 27650],
        ['MKT-013', 'Video Ad Campaign', 16800, 1120000, 25760, 515, 30900],
        ['MKT-014', 'Podcast Sponsorship', 11000, 640000, 12800, 384, 26880],
        ['MKT-015', 'Search Engine Marketing', 28500, 1900000, 57000, 1710, 136800],
        ['MKT-016', 'Display Retargeting', 6300, 450000, 10350, 207, 12420],
        ['MKT-017', 'Content Marketing Hub', 4800, 320000, 8000, 320, 22400],
        ['MKT-018', 'Affiliate Network Push', 13700, 870000, 19140, 574, 40180],
        ['MKT-019', 'Webinar Lead Gen', 9100, 560000, 14560, 582, 43650],
        ['MKT-020', 'Mobile App Install', 19500, 1350000, 33750, 675, 40500],
    ]

    for r, row_data in enumerate(campaigns, 2):
        # Columns A-G: direct values
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
        # Column H: CTR = Clicks / Impressions
        ws.cell(row=r, column=8, value=f'=E{r}/D{r}')
        # Column I: Conversion Rate = Conversions / Clicks
        ws.cell(row=r, column=9, value=f'=F{r}/E{r}')
        # Column J: ROAS = Revenue / Spend
        ws.cell(row=r, column=10, value=f'=G{r}/C{r}')

    # Row 22 is intentionally left EMPTY (task asks agent to add summary row)
    # NO formatting applied (task asks agent to format)
    # NO bold on row 1 (task asks agent to bold headers)
    # NO borders (task asks agent to add borders)
    # NO conditional formatting (task asks agent to add it)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
