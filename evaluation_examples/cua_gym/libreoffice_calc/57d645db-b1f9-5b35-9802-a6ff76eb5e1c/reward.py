"""
Reward Script: Marketing campaign performance summary
Task ID: calc_gsd_048
Domain: libreoffice_calc
Scoring:
  Component 1: Row 22 summary row with correct formulas (0.25)
  Component 2: Row 22 bold formatting (0.10)
  Component 3: Currency formatting on C2:C21 and G2:G21 (0.15)
  Component 4: Percentage formatting on H2:H21 and I2:I21 (0.10)
  Component 5: Number formatting on J2:J21 (0.05)
  Component 6: Row 1 bold (0.10)
  Component 7: Borders on A1:J22 (0.10)
  Component 8: Conditional formatting on J2:J21 with gold background (0.15)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_048'


def persist_app_state(domain):
    """Try to save any unsaved changes in LibreOffice."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Campaigns' not in wb.sheetnames:
        print("CRITICAL: 'Campaigns' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Campaigns']

    # Component 1: Row 22 summary row with correct formulas (0.25 points)
    try:
        a22 = ws['A22'].value
        c22 = ws['C22'].value
        g22 = ws['G22'].value
        h22 = ws['H22'].value
        i22 = ws['I22'].value
        j22 = ws['J22'].value

        comp1_score = 0.0
        # Check A22 has "Summary" label
        if a22 and str(a22).strip().lower() == 'summary':
            comp1_score += 0.05
        else:
            print(f"FAIL: Component 1a — A22 expected 'Summary', found: {a22}")

        # Check C22 has SUM formula
        if c22 and isinstance(c22, str) and 'SUM' in c22.upper() and 'C2' in c22.upper():
            comp1_score += 0.04
        else:
            print(f"FAIL: Component 1b — C22 expected SUM formula, found: {c22}")

        # Check G22 has SUM formula
        if g22 and isinstance(g22, str) and 'SUM' in g22.upper() and 'G2' in g22.upper():
            comp1_score += 0.04
        else:
            print(f"FAIL: Component 1c — G22 expected SUM formula, found: {g22}")

        # Check H22 has AVERAGE formula
        if h22 and isinstance(h22, str) and 'AVERAGE' in h22.upper() and 'H2' in h22.upper():
            comp1_score += 0.04
        else:
            print(f"FAIL: Component 1d — H22 expected AVERAGE formula, found: {h22}")

        # Check I22 has AVERAGE formula
        if i22 and isinstance(i22, str) and 'AVERAGE' in i22.upper() and 'I2' in i22.upper():
            comp1_score += 0.04
        else:
            print(f"FAIL: Component 1e — I22 expected AVERAGE formula, found: {i22}")

        # Check J22 has AVERAGE formula
        if j22 and isinstance(j22, str) and 'AVERAGE' in j22.upper() and 'J2' in j22.upper():
            comp1_score += 0.04
        else:
            print(f"FAIL: Component 1f — J22 expected AVERAGE formula, found: {j22}")

        if comp1_score > 0:
            print(f"PASS: Component 1 — Summary row formulas ({comp1_score} pts)")
        total_score += comp1_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Row 22 bold formatting (0.10 points)
    try:
        row22_bold_count = 0
        for c in range(1, 11):
            if ws.cell(row=22, column=c).font.bold:
                row22_bold_count += 1
        # Need at least the cells that have values (A, C, G, H, I, J = 6 cells) to be bold
        # But task says make row 22 bold, so check all 10
        if row22_bold_count >= 6:
            print(f"PASS: Component 2 — Row 22 bold ({row22_bold_count}/10 cells bold) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 — Row 22 bold: only {row22_bold_count}/10 cells are bold")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Currency formatting on C2:C21 and G2:G21 (0.15 points)
    try:
        currency_ok = 0
        total_cells = 40  # 20 rows * 2 columns
        for row in range(2, 22):
            for col_letter in ['C', 'G']:
                col = 3 if col_letter == 'C' else 7
                nf = ws.cell(row=row, column=col).number_format
                # Accept any format containing '$' as currency
                if nf and '$' in str(nf):
                    currency_ok += 1
        ratio = currency_ok / total_cells
        if ratio >= 0.9:
            print(f"PASS: Component 3 — Currency format on C,G columns ({currency_ok}/{total_cells}) (0.15 pts)")
            total_score += 0.15
        elif ratio >= 0.5:
            partial = round(0.15 * ratio, 3)
            print(f"PARTIAL: Component 3 — Currency format ({currency_ok}/{total_cells}) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Currency format: only {currency_ok}/{total_cells} cells formatted")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Percentage formatting on H2:H21 and I2:I21 (0.10 points)
    try:
        pct_ok = 0
        total_cells = 40  # 20 rows * 2 columns
        for row in range(2, 22):
            for col in [8, 9]:  # H, I
                nf = ws.cell(row=row, column=col).number_format
                if nf and '%' in str(nf):
                    pct_ok += 1
        ratio = pct_ok / total_cells
        if ratio >= 0.9:
            print(f"PASS: Component 4 — Percentage format on H,I columns ({pct_ok}/{total_cells}) (0.10 pts)")
            total_score += 0.10
        elif ratio >= 0.5:
            partial = round(0.10 * ratio, 3)
            print(f"PARTIAL: Component 4 — Percentage format ({pct_ok}/{total_cells}) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Percentage format: only {pct_ok}/{total_cells} cells formatted")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Number formatting on J2:J21 with 2 decimals (0.05 points)
    try:
        num_ok = 0
        for row in range(2, 22):
            nf = ws.cell(row=row, column=10).number_format  # J column
            # Accept formats with decimal places that are NOT percentage
            if nf and nf != 'General' and '%' not in str(nf):
                # Check it has decimal formatting (contains '0.00' or '#.##' etc.)
                if '.' in str(nf):
                    num_ok += 1
        if num_ok >= 18:  # 90% of 20
            print(f"PASS: Component 5 — Number format on J column ({num_ok}/20) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — Number format on J column: only {num_ok}/20 cells formatted")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Row 1 bold (0.10 points)
    try:
        row1_bold_count = 0
        for c in range(1, 11):
            if ws.cell(row=1, column=c).font.bold:
                row1_bold_count += 1
        if row1_bold_count >= 8:  # at least 8 of 10 headers bold
            print(f"PASS: Component 6 — Row 1 bold ({row1_bold_count}/10) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 6 — Row 1 bold: only {row1_bold_count}/10 cells are bold")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: Borders on A1:J22 (0.10 points)
    try:
        bordered_count = 0
        total_border_cells = 22 * 10  # 22 rows, 10 columns
        for row in range(1, 23):
            for col in range(1, 11):
                cell = ws.cell(row=row, column=col)
                b = cell.border
                # Check if at least some borders are set
                has_border = any([
                    b.left.style is not None,
                    b.right.style is not None,
                    b.top.style is not None,
                    b.bottom.style is not None
                ])
                if has_border:
                    bordered_count += 1
        ratio = bordered_count / total_border_cells
        if ratio >= 0.8:
            print(f"PASS: Component 7 — Borders ({bordered_count}/{total_border_cells} cells) (0.10 pts)")
            total_score += 0.10
        elif ratio >= 0.3:
            partial = round(0.10 * ratio, 3)
            print(f"PARTIAL: Component 7 — Borders ({bordered_count}/{total_border_cells}) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 7 — Borders: only {bordered_count}/{total_border_cells} cells have borders")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    # Component 8: Conditional formatting on J2:J21 with gold background (0.15 points)
    try:
        cf_list = list(ws.conditional_formatting)
        found_gold_cf = False
        for cf in cf_list:
            cf_range_str = str(cf).upper()
            # Check if the CF covers J2:J21 range (column J)
            if 'J' in cf_range_str:
                for rule in cf.rules:
                    # Check for gold color in the differential style
                    if rule.dxf and rule.dxf.fill:
                        fg = rule.dxf.fill.fgColor
                        if fg and fg.rgb:
                            rgb_str = str(fg.rgb).upper()
                            # Gold is #FFD700 -> ARGB FFFFD700
                            if 'FFD700' in rgb_str:
                                found_gold_cf = True
                                break
            if found_gold_cf:
                break

        if found_gold_cf:
            print(f"PASS: Component 8 — Conditional formatting with gold background on J column (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 8 — No conditional formatting with gold (#FFD700) background found on J column")
            print(f"  Found {len(cf_list)} CF rule(s)")
    except Exception as e:
        print(f"ERROR: Component 8 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
