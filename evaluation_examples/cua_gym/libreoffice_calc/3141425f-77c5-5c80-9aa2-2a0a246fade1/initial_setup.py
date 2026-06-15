"""
Initial Setup: Fix percentage display - cells show 2500% instead of 25%
Task ID: calc_tbl_092
Domain: libreoffice_calc

Creates a spreadsheet where column C has whole-number values (25, 30, etc.)
formatted as percentages, causing them to display as 2500%, 3000%, etc.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_092'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Employee Performance ---
    ws = wb.active
    ws.title = 'Employee Performance'

    # Headers
    headers = ['Employee Name', 'Department', 'Performance Rating', 'Annual Bonus ($)']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows - 19 employees
    employees = [
        ['Sarah Chen', 'Engineering', 85, 12500],
        ['Marcus Johnson', 'Marketing', 72, 9800],
        ['Priya Patel', 'Engineering', 91, 15200],
        ['James O\'Brien', 'Sales', 68, 8400],
        ['Yuki Tanaka', 'Product', 78, 10600],
        ['Elena Rodriguez', 'Marketing', 65, 7900],
        ['David Kim', 'Engineering', 88, 13800],
        ['Aisha Mohammed', 'Sales', 93, 16000],
        ['Thomas Wright', 'Finance', 71, 9500],
        ['Lena Johansson', 'Product', 82, 11200],
        ['Roberto Diaz', 'Sales', 76, 10100],
        ['Megan Foster', 'Finance', 69, 8600],
        ['Wei Zhang', 'Engineering', 95, 17500],
        ['Olivia Bennett', 'Marketing', 74, 9900],
        ['Rajesh Kumar', 'Product', 80, 10800],
        ['Sophie Martin', 'Finance', 67, 8100],
        ['Hassan Ali', 'Sales', 87, 13200],
        ['Catherine Lee', 'Engineering', 90, 14800],
        ['Michael Torres', 'Marketing', 73, 9700],
    ]

    data_align = Alignment(horizontal='center', vertical='center')

    for r, row_data in enumerate(employees, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c >= 2:
                cell.alignment = data_align
            # Column C: format as percentage (BUT values are whole numbers like 85, 72...)
            # This causes display of 8500%, 7200%, etc. - the bug the task asks to fix
            if c == 3:
                cell.number_format = '0%'

    # Set column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18

    # Freeze header row
    ws.freeze_panes = 'A2'

    # --- Sheet 2: Department Summary ---
    ws2 = wb.create_sheet('Department Summary')
    dept_headers = ['Department', 'Headcount', 'Avg Performance', 'Total Bonus Budget']
    for col, h in enumerate(dept_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    dept_data = [
        ['Engineering', 5, 90, 73800],
        ['Marketing', 4, 71, 37300],
        ['Sales', 4, 81, 47700],
        ['Product', 3, 80, 32600],
        ['Finance', 3, 69, 26200],
    ]
    for r, row_data in enumerate(dept_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c >= 2:
                cell.alignment = data_align

    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 20

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
