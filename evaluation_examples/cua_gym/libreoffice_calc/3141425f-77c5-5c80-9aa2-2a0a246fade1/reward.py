"""
Reward Script: Fix percentage format display in column C
Task ID: calc_tbl_092
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): Values in C2:C20 are decimal floats (0 < v < 1), not integers
  Component 2 (0.3): All 19 values match expected decimals (original/100) exactly
  Component 3 (0.2): Percentage number format ('0%' or similar) preserved on C2:C20
"""

import os

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_092'

# Expected values: initial integers divided by 100
EXPECTED_VALUES = {
    2: 0.85, 3: 0.72, 4: 0.91, 5: 0.68, 6: 0.78,
    7: 0.65, 8: 0.88, 9: 0.93, 10: 0.71, 11: 0.82,
    12: 0.76, 13: 0.69, 14: 0.95, 15: 0.74, 16: 0.80,
    17: 0.67, 18: 0.87, 19: 0.90, 20: 0.73,
}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Component 1: Values in C2:C20 are decimal floats in (0, 1) range (0.5 points)
    # This checks that values have been converted from integers (like 85) to decimals (like 0.85)
    # On initial_env, values are integers >= 1, so this FAILS. On golden_env, values are < 1, so this PASSES.
    try:
        decimal_count = 0
        total_cells = 19  # C2 through C20
        for row in range(2, 21):
            val = ws.cell(row=row, column=3).value
            if val is not None and isinstance(val, (int, float)):
                fval = float(val)
                if 0 < fval < 1:
                    decimal_count += 1

        if decimal_count == total_cells:
            print(f"PASS: Component 1 — All {total_cells} values in C2:C20 are decimals in (0,1) range (0.5 pts)")
            total_score += 0.5
        elif decimal_count > 0:
            partial = 0.5 * (decimal_count / total_cells)
            print(f"PARTIAL: Component 1 — {decimal_count}/{total_cells} values are decimals (0,1). Partial: {partial:.3f} pts")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — 0/{total_cells} values are decimals in (0,1) range. Values are still integers or out of range.")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All 19 values match expected decimals exactly (0.3 points)
    # Verifies precise conversion: each original integer / 100 = expected decimal
    # On initial_env, values are integers so this FAILS. On golden_env, values match.
    try:
        match_count = 0
        for row, expected in EXPECTED_VALUES.items():
            val = ws.cell(row=row, column=3).value
            if val is not None:
                try:
                    fval = float(val)
                    if abs(fval - expected) < 0.001:
                        match_count += 1
                    else:
                        print(f"  C{row}: expected {expected}, got {fval}")
                except (ValueError, TypeError):
                    print(f"  C{row}: non-numeric value {val!r}")

        if match_count == len(EXPECTED_VALUES):
            print(f"PASS: Component 2 — All {len(EXPECTED_VALUES)} values match expected decimals exactly (0.3 pts)")
            total_score += 0.3
        elif match_count > 0:
            partial = 0.3 * (match_count / len(EXPECTED_VALUES))
            print(f"PARTIAL: Component 2 — {match_count}/{len(EXPECTED_VALUES)} values match. Partial: {partial:.3f} pts")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — 0/{len(EXPECTED_VALUES)} values match expected decimals")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Percentage number format preserved on C2:C20 (0.2 points)
    # The format should still be percentage-style (contains '%').
    # On initial_env, format is '0%' BUT values are integers — so Component 1 and 2 gate this.
    # We only award these points if Component 1 passed (values are decimals).
    # This ensures initial_env scores 0.0 for this component.
    try:
        pct_format_count = 0
        for row in range(2, 21):
            fmt = ws.cell(row=row, column=3).number_format
            if fmt and '%' in str(fmt):
                pct_format_count += 1

        # Gate: only award if values are actually decimals (Component 1 condition)
        vals_are_decimal = all(
            isinstance(ws.cell(row=r, column=3).value, (int, float)) and 0 < float(ws.cell(row=r, column=3).value) < 1
            for r in range(2, 21)
            if ws.cell(row=r, column=3).value is not None
        )

        if pct_format_count == 19 and vals_are_decimal:
            print(f"PASS: Component 3 — All 19 cells retain percentage format AND values are decimal (0.2 pts)")
            total_score += 0.2
        elif pct_format_count == 19 and not vals_are_decimal:
            print(f"FAIL: Component 3 — Format is percentage but values are not decimal yet (gated)")
        else:
            print(f"FAIL: Component 3 — Only {pct_format_count}/19 cells have percentage format")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
