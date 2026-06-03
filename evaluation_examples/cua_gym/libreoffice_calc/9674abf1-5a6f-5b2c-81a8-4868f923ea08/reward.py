"""
Reward Script: Add dropdown list data validation to C2:C51 on 'Orders' sheet
Task ID: calc_gg3_007
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): Data validation exists on C2:C51 with type 'list'
  Component 2 (0.25): Validation formula contains all 5 required status values
  Component 3 (0.20): Input message title 'Order Status' and message correct
  Component 4 (0.20): All C2:C51 values are valid standardized entries
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_007'

VALID_STATUSES = {'Pending', 'Processing', 'Shipped', 'Delivered', 'Cancelled'}


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Orders' sheet must exist
    if 'Orders' not in wb.sheetnames:
        print(f"CRITICAL: 'Orders' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Orders']

    # Component 1: Data validation exists on C2:C51 with type 'list' (0.35 points)
    try:
        dvs = ws.data_validations.dataValidation
        matching_dv = None
        for dv in dvs:
            if dv.type == 'list':
                # Check if the validation covers C2:C51
                sqref_str = str(dv.sqref)
                # The sqref could be "C2:C51" or contain it
                if 'C2' in sqref_str and 'C51' in sqref_str:
                    matching_dv = dv
                    break

        if matching_dv is not None:
            print(f"PASS: Component 1 — List data validation found on {matching_dv.sqref} (0.35 pts)")
            total_score += 0.35
        else:
            # Check if there's any list validation at all
            list_dvs = [dv for dv in dvs if dv.type == 'list']
            if list_dvs:
                print(f"FAIL: Component 1 — List validation found but not covering C2:C51. Found: {[str(dv.sqref) for dv in list_dvs]}")
            else:
                print(f"FAIL: Component 1 — No list data validation found. Total validations: {len(dvs)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        matching_dv = None

    # Component 2: Validation formula contains all 5 required status values (0.25 points)
    try:
        if matching_dv is not None:
            formula = matching_dv.formula1 or ''
            # Remove surrounding quotes if present
            formula_clean = formula.strip('"').strip("'")
            # Parse comma-separated values
            formula_values = {v.strip() for v in formula_clean.split(',')}

            missing = VALID_STATUSES - formula_values
            extra = formula_values - VALID_STATUSES

            if not missing:
                print(f"PASS: Component 2 — All 5 status values present in validation formula (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 2 — Missing values: {missing}. Found: {formula_values}")
        else:
            print(f"FAIL: Component 2 — No matching data validation to check formula")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Input message title 'Order Status' and message correct (0.20 points)
    try:
        if matching_dv is not None:
            title_ok = (matching_dv.promptTitle or '').strip() == 'Order Status'
            msg_ok = (matching_dv.prompt or '').strip() == 'Select a status from the list.'

            if title_ok and msg_ok:
                print(f"PASS: Component 3 — Input message title and message correct (0.20 pts)")
                total_score += 0.20
            elif title_ok:
                print(f"FAIL: Component 3 — Title correct but message wrong. Found: '{matching_dv.prompt}'")
                total_score += 0.10  # partial credit for title
            elif msg_ok:
                print(f"FAIL: Component 3 — Message correct but title wrong. Found: '{matching_dv.promptTitle}'")
                total_score += 0.10  # partial credit for message
            else:
                print(f"FAIL: Component 3 — Both title and message wrong. Title: '{matching_dv.promptTitle}', Message: '{matching_dv.prompt}'")
        else:
            print(f"FAIL: Component 3 — No matching data validation to check input message")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: All C2:C51 values are valid standardized entries (0.20 points)
    try:
        valid_count = 0
        total_cells = 50
        for r in range(2, 52):
            val = ws.cell(row=r, column=3).value
            if val in VALID_STATUSES:
                valid_count += 1

        if valid_count == total_cells:
            print(f"PASS: Component 4 — All {total_cells} cells have valid status values (0.20 pts)")
            total_score += 0.20
        elif valid_count >= 40:
            partial = round(0.20 * (valid_count / total_cells), 2)
            print(f"PARTIAL: Component 4 — {valid_count}/{total_cells} cells valid (partial: {partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {valid_count}/{total_cells} cells have valid status values")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
