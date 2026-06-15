"""
Initial Setup: Add dropdown data validation to order status column
Task ID: calc_gg3_007
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Orders Sheet ---
    ws = wb.active
    ws.title = 'Orders'

    # Headers
    headers = ['Order ID', 'Customer', 'Status', 'Order Date', 'Amount']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic customer names
    customers = [
        'Sarah Chen', 'Marcus Johnson', 'Elena Rodriguez', 'James Kim',
        'Priya Patel', 'David Thompson', 'Aisha Mohammed', 'Ryan O\'Brien',
        'Yuki Tanaka', 'Isabella Costa', 'Ahmed Hassan', 'Emma Larsson',
        'Carlos Mendoza', 'Fatima Al-Rashid', 'Liam McCarthy',
        'Sophia Rossi', 'Noah Williams', 'Mia Nakamura', 'Ethan Brown',
        'Olivia Taylor', 'Lucas Fischer', 'Ava Dubois', 'Benjamin Lee',
        'Charlotte Moore', 'Alexander Petrov', 'Hannah Schwartz',
        'Daniel Garcia', 'Grace Nguyen', 'Samuel Clark', 'Lily Chang',
        'Thomas Anderson', 'Zoe Palmer', 'William Chen', 'Natalie Burke',
        'Henry Wright', 'Chloe Martin', 'Oscar Rivera', 'Amelia Cooper',
        'Jack Robinson', 'Emily Watson', 'Leo Hoffman', 'Maya Singh',
        'Oliver James', 'Aria Fernandez', 'Sebastian Park', 'Victoria Adams',
        'Felix Turner', 'Stella Morgan', 'Max Bennett', 'Ruby Scott',
    ]

    # Inconsistent free-text statuses (the problem the task is meant to fix)
    inconsistent_statuses = [
        'pending', 'Pending', 'PENDING', 'pend',
        'shipped', 'Shipped', 'SHIPPED', 'ship', 'shiped',
        'delivered', 'Delivered', 'DELIVERED', 'deliver', 'delievered',
        'processing', 'Processing', 'PROCESSING', 'in process', 'proc',
        'cancelled', 'Cancelled', 'CANCELLED', 'cancel', 'canceled',
    ]

    random.seed(42)

    for i in range(50):
        row = i + 2
        # Order ID: ORD-2025-XXXX
        order_id = f'ORD-2025-{1001 + i}'
        ws.cell(row=row, column=1, value=order_id)

        # Customer
        ws.cell(row=row, column=2, value=customers[i])

        # Inconsistent status (free-text)
        status = random.choice(inconsistent_statuses)
        ws.cell(row=row, column=3, value=status)

        # Order Date: dates in 2025
        month = random.randint(1, 12)
        day = random.randint(1, 28)
        ws.cell(row=row, column=4, value=f'2025-{month:02d}-{day:02d}')

        # Amount: between $15 and $2500
        amount = round(random.uniform(15, 2500), 2)
        ws.cell(row=row, column=5, value=amount)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12

    # Number format for Amount column
    for row in range(2, 52):
        ws.cell(row=row, column=5).number_format = '$#,##0.00'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
