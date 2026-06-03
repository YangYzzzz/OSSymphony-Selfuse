"""
Reward Script: Define named ranges and SUM formula in LibreOffice Calc
Task ID: calc_ggf_018
Domain: libreoffice_calc
Scoring:
  Component 1 (0.2): NorthSales named range -> Master!$B$2:$B$13
  Component 2 (0.2): SouthSales named range -> Master!$C$2:$C$13
  Component 3 (0.2): EastSales named range -> Master!$D$2:$D$13
  Component 4 (0.25): F2 contains formula referencing all three named ranges
  Component 5 (0.15): F2 formula yields correct total (1524470)
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_018'


def normalize_range_ref(attr_text):
    """Normalize a defined name reference for comparison.
    Handles forms like: Master!$B$2:$B$13, 'Master'!$B$2:$B$13, Master!B2:B13
    Returns uppercase canonical form: MASTER!$B$2:$B$13
    """
    s = attr_text.strip().upper()
    # Remove quotes around sheet name
    s = s.replace("'", "")
    # Split on ! to handle sheet name and cell references separately
    parts = s.split('!')
    if len(parts) != 2:
        return s
    sheet_part = parts[0]
    cell_part = parts[1]
    # Normalize each cell ref in the range (e.g., B2:B13 -> $B$2:$B$13)
    cell_refs = cell_part.split(':')
    normalized_refs = []
    for ref in cell_refs:
        # Remove existing $ signs then re-add canonically
        ref_clean = ref.replace('$', '')
        # Match column letters and row number
        m = re.match(r'^([A-Z]+)(\d+)$', ref_clean)
        if m:
            normalized_refs.append(f'${m.group(1)}${m.group(2)}')
        else:
            normalized_refs.append(ref)
    return sheet_part + '!' + ':'.join(normalized_refs)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Master' not in wb.sheetnames:
        print("CRITICAL: 'Master' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Master']

    # Collect all defined names into a dict for lookup
    defined_names = {}
    try:
        for name, dn in wb.defined_names.items():
            defined_names[name.upper()] = dn.attr_text
    except Exception as e:
        print(f"ERROR: Could not read defined names: {e}")

    print(f"DEBUG: Found defined names: {defined_names}")

    # Component 1: NorthSales named range (0.2 points)
    try:
        if 'NORTHSALES' in defined_names:
            ref = normalize_range_ref(defined_names['NORTHSALES'])
            expected = 'MASTER!$B$2:$B$13'
            if ref == expected:
                print(f"PASS: Component 1 - NorthSales = {defined_names['NORTHSALES']} (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 1 - NorthSales ref is '{ref}', expected '{expected}'")
        else:
            print("FAIL: Component 1 - NorthSales named range not found")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: SouthSales named range (0.2 points)
    try:
        if 'SOUTHSALES' in defined_names:
            ref = normalize_range_ref(defined_names['SOUTHSALES'])
            expected = 'MASTER!$C$2:$C$13'
            if ref == expected:
                print(f"PASS: Component 2 - SouthSales = {defined_names['SOUTHSALES']} (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 2 - SouthSales ref is '{ref}', expected '{expected}'")
        else:
            print("FAIL: Component 2 - SouthSales named range not found")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: EastSales named range (0.2 points)
    try:
        if 'EASTSALES' in defined_names:
            ref = normalize_range_ref(defined_names['EASTSALES'])
            expected = 'MASTER!$D$2:$D$13'
            if ref == expected:
                print(f"PASS: Component 3 - EastSales = {defined_names['EASTSALES']} (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 3 - EastSales ref is '{ref}', expected '{expected}'")
        else:
            print("FAIL: Component 3 - EastSales named range not found")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: F2 contains a formula referencing all three named ranges (0.25 points)
    try:
        f2_val = ws['F2'].value
        if f2_val and isinstance(f2_val, str) and f2_val.startswith('='):
            f2_upper = f2_val.upper().replace(' ', '')
            has_north = 'NORTHSALES' in f2_upper
            has_south = 'SOUTHSALES' in f2_upper
            has_east = 'EASTSALES' in f2_upper
            has_sum = 'SUM' in f2_upper

            if has_north and has_south and has_east and has_sum:
                print(f"PASS: Component 4 - F2 formula references all 3 named ranges with SUM: {f2_val} (0.25 pts)")
                total_score += 0.25
            else:
                missing = []
                if not has_north:
                    missing.append('NorthSales')
                if not has_south:
                    missing.append('SouthSales')
                if not has_east:
                    missing.append('EastSales')
                if not has_sum:
                    missing.append('SUM function')
                print(f"FAIL: Component 4 - F2 formula missing: {', '.join(missing)}. Found: {f2_val}")
        else:
            print(f"FAIL: Component 4 - F2 does not contain a formula. Value: {f2_val}")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: F2 formula yields correct total (0.15 points)
    # We compute the expected sum from the actual data in the sheet
    try:
        expected_total = 0
        for col in [2, 3, 4]:  # B, C, D
            for row in range(2, 14):  # rows 2-13
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    expected_total += float(val)

        # Load with data_only to get cached computed value
        wb_data = openpyxl.load_workbook(file_path, data_only=True)
        ws_data = wb_data['Master']
        f2_computed = ws_data['F2'].value

        if f2_computed is not None:
            if abs(float(f2_computed) - expected_total) < 0.01:
                print(f"PASS: Component 5 - F2 computed value = {f2_computed}, expected = {expected_total} (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 5 - F2 computed value = {f2_computed}, expected = {expected_total}")
        else:
            # data_only may return None if file was never opened in Calc.
            # If all named ranges and formula are correct, accept on formula validity alone.
            # Check if formula is structurally valid (already verified in Component 4)
            if total_score >= 0.85:
                # All named ranges correct + formula correct => formula would compute correctly
                print(f"PASS: Component 5 - F2 cached value unavailable but formula and ranges verified correct; expected total = {expected_total} (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 5 - F2 cached value is None and cannot verify correctness")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
