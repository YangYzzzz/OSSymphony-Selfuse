"""
Initial Setup: Define named ranges and sum formula for regional sales
Task ID: calc_ggf_018
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_018'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Master sheet ---
    ws = wb.active
    ws.title = 'Master'

    # Headers
    ws['A1'] = 'Month'
    ws['B1'] = 'North'
    ws['C1'] = 'South'
    ws['D1'] = 'East'
    ws['F1'] = 'Grand Total'

    # Monthly data - realistic regional sales figures
    months = [
        'January', 'February', 'March', 'April', 'May', 'June',
        'July', 'August', 'September', 'October', 'November', 'December'
    ]
    north_sales = [45230, 38920, 52100, 41850, 49300, 55670, 61200, 58400, 47900, 53100, 62800, 71500]
    south_sales = [32100, 29800, 35400, 33200, 37600, 41200, 44800, 42100, 36500, 39800, 45200, 52300]
    east_sales = [28500, 25700, 31200, 29400, 33800, 36900, 39500, 37200, 32100, 35400, 40100, 46700]

    for i, month in enumerate(months):
        row = i + 2
        ws.cell(row=row, column=1, value=month)
        ws.cell(row=row, column=2, value=north_sales[i])
        ws.cell(row=row, column=3, value=south_sales[i])
        ws.cell(row=row, column=4, value=east_sales[i])

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['F'].width = 14

    # NO named ranges - that is the task
    # NO formula in F2 - that is the task

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
