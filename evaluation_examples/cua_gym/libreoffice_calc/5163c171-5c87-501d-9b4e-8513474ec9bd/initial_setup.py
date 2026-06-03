"""
Initial Setup: HR Severance Calculation
Task ID: calc_hr_severance_calculation_064
Domain: libreoffice_calc

Creates the RIF List spreadsheet with employee data (columns A-E filled,
columns F-H empty — those are what the agent must calculate).
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_severance_calculation_064'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'RIF List'

    # --- Headers (Row 1) ---
    headers = [
        'Emp ID', 'Name', 'Hire Date', 'Separation Date',
        'Annual Salary', 'Years of Service', 'Severance Weeks', 'Severance Amount'
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # --- Employee data (rows 2-34): columns A-E filled, F-H empty ---
    # 33 employees being separated in a reduction-in-force
    employees = [
        ('E1001', 'Sarah Chen',        date(2012, 3, 14),  date(2025, 6, 30),  92500),
        ('E1002', 'Marcus Johnson',     date(2015, 7, 1),   date(2025, 6, 30),  78000),
        ('E1003', 'Linda Okafor',       date(2008, 11, 20), date(2025, 6, 30), 105000),
        ('E1004', 'James Whitfield',    date(2019, 2, 10),  date(2025, 6, 30),  67500),
        ('E1005', 'Priya Nair',         date(2010, 5, 3),   date(2025, 6, 30),  88000),
        ('E1006', 'Derek Santos',       date(2016, 9, 15),  date(2025, 6, 30),  74000),
        ('E1007', 'Angela Reyes',       date(2007, 1, 22),  date(2025, 6, 30), 112000),
        ('E1008', 'Thomas Nguyen',      date(2021, 4, 5),   date(2025, 6, 30),  60000),
        ('E1009', 'Karen Fitzgerald',   date(2013, 8, 30),  date(2025, 6, 30),  83000),
        ('E1010', 'Robert Ibarra',      date(2009, 6, 17),  date(2025, 6, 30),  97500),
        ('E1011', 'Michelle Larsson',   date(2018, 3, 11),  date(2025, 6, 30),  71000),
        ('E1012', 'David Osei',         date(2014, 12, 2),  date(2025, 6, 30),  79500),
        ('E1013', 'Rachel Kowalski',    date(2011, 10, 28), date(2025, 6, 30),  86000),
        ('E1014', 'Steven Yamamoto',    date(2006, 4, 8),   date(2025, 6, 30), 118000),
        ('E1015', 'Natalie Adeyemi',    date(2020, 1, 13),  date(2025, 6, 30),  64000),
        ('E1016', 'Brian Castellano',   date(2017, 6, 26),  date(2025, 6, 30),  73000),
        ('E1017', 'Cynthia Park',       date(2012, 9, 7),   date(2025, 6, 30),  89000),
        ('E1018', 'Anthony Moreau',     date(2005, 2, 19),  date(2025, 6, 30), 125000),
        ('E1019', 'Stephanie Obi',      date(2022, 7, 1),   date(2025, 6, 30),  57000),
        ('E1020', 'Kevin Thornton',     date(2010, 11, 3),  date(2025, 6, 30),  93000),
        ('E1021', 'Jessica Alvarez',    date(2015, 3, 24),  date(2025, 6, 30),  77000),
        ('E1022', 'Raymond Chukwu',     date(2008, 8, 16),  date(2025, 6, 30), 101000),
        ('E1023', 'Pamela Sorensen',    date(2019, 5, 9),   date(2025, 6, 30),  68500),
        ('E1024', 'Gregory Watkins',    date(2013, 1, 31),  date(2025, 6, 30),  84000),
        ('E1025', 'Laura Mensah',       date(2016, 10, 18), date(2025, 6, 30),  76000),
        ('E1026', 'Patrick Huang',      date(2007, 7, 4),   date(2025, 6, 30), 109000),
        ('E1027', 'Sandra Eriksson',    date(2021, 9, 27),  date(2025, 6, 30),  62000),
        ('E1028', 'Michael Dembele',    date(2011, 4, 12),  date(2025, 6, 30),  90500),
        ('E1029', 'Tiffany Goldstein',  date(2014, 6, 5),   date(2025, 6, 30),  80000),
        ('E1030', 'Charles Svensson',   date(2009, 12, 23), date(2025, 6, 30),  98000),
        ('E1031', 'Denise Achebe',      date(2018, 2, 14),  date(2025, 6, 30),  72500),
        ('E1032', 'Victor Nakamura',    date(2006, 9, 1),   date(2025, 6, 30), 115000),
        ('E1033', 'Monique Diallo',     date(2023, 1, 16),  date(2025, 6, 30),  55000),
    ]

    for r, (emp_id, name, hire_date, sep_date, salary) in enumerate(employees, 2):
        ws.cell(row=r, column=1, value=emp_id)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=hire_date)
        ws.cell(row=r, column=4, value=sep_date)
        ws.cell(row=r, column=5, value=salary)
        # Columns F (6), G (7), H (8) intentionally left empty

    # Format date columns
    for r in range(2, 35):
        ws.cell(row=r, column=3).number_format = 'yyyy-mm-dd'
        ws.cell(row=r, column=4).number_format = 'yyyy-mm-dd'

    # Format salary column
    for r in range(2, 35):
        ws.cell(row=r, column=5).number_format = '$#,##0.00'

    # Column widths for readability
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 17
    ws.column_dimensions['H'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: RIF List')
    print(f'  Rows: 1 header + 33 employee rows (2-34)')
    print(f'  Columns F/G/H: EMPTY (to be filled by agent)')


create_initial()
