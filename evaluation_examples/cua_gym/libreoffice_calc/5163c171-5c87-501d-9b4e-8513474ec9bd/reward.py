"""
Reward Script: Calculate severance packages for reduction-in-force list
Task ID: calc_hr_severance_calculation_064
Domain: libreoffice_calc
Scoring:
  - Component 1: DATEDIF years-of-service formulas in F2:F34 with integer format (0.25)
  - Component 2: MIN/MAX severance weeks formulas in G2:G34 with integer format (0.25)
  - Component 3: Severance amount formulas in H2:H34 with currency format $#,##0.00 (0.25)
  - Component 4: Total row - label in G35, SUM formula in H35, both bold, H35 currency format (0.25)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_severance_calculation_064'


def normalize_formula(formula):
    """Normalize a formula string for comparison: uppercase, remove spaces."""
    if not isinstance(formula, str):
        return ''
    return formula.upper().replace(' ', '').replace('"', "'")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that 'RIF List' sheet exists (precondition gate)
    if 'RIF List' not in wb.sheetnames:
        print("CRITICAL: Sheet 'RIF List' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['RIF List']

    # Component 1: DATEDIF years-of-service formulas in F2:F34 (0.25 points)
    # Verifies the agent added =DATEDIF(CX,DX,"Y") to each row — absent in initial file
    try:
        f_formula_count = 0
        f_format_count = 0
        f_expected_formula_prefix = '=DATEDIF('
        for row in range(2, 35):
            cell = ws.cell(row=row, column=6)
            val = cell.value
            nf = cell.number_format

            # Check for DATEDIF formula
            if (isinstance(val, str) and
                    val.upper().startswith('=DATEDIF(') and
                    '"Y"' in val.upper().replace("'Y'", '"Y"')):
                f_formula_count += 1

            # Check for integer number format (0 or General/integer-like)
            if nf in ('0', '0.00', 'General', '#,##0'):
                f_format_count += 1

        if f_formula_count == 33:
            print(f"PASS: Component 1 — All 33 rows have DATEDIF formula in F column ({f_formula_count}/33) (0.25 pts)")
            total_score += 0.25
        elif f_formula_count >= 25:
            print(f"PARTIAL: Component 1 — {f_formula_count}/33 rows have DATEDIF formula in F column (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 — Expected DATEDIF formulas in F2:F34, found {f_formula_count}/33 rows")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: MIN/MAX severance weeks formulas in G2:G34 (0.25 points)
    # Verifies the agent added =MIN(MAX(FX*2,4),26) to each row — absent in initial file
    try:
        g_formula_count = 0
        for row in range(2, 35):
            cell = ws.cell(row=row, column=7)
            val = cell.value
            nf = cell.number_format

            # Check for MIN(MAX(...)) formula with correct logic
            if isinstance(val, str):
                val_up = val.upper().replace(' ', '')
                if '=MIN(' in val_up and 'MAX(' in val_up and '*2' in val_up:
                    g_formula_count += 1

        if g_formula_count == 33:
            print(f"PASS: Component 2 — All 33 rows have MIN(MAX()) formula in G column ({g_formula_count}/33) (0.25 pts)")
            total_score += 0.25
        elif g_formula_count >= 25:
            print(f"PARTIAL: Component 2 — {g_formula_count}/33 rows have MIN/MAX formula in G column (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 — Expected MIN(MAX()) formulas in G2:G34, found {g_formula_count}/33 rows")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Severance amount formulas in H2:H34 with $#,##0.00 format (0.25 points)
    # Verifies agent added =(EX/52)*GX with currency format — absent in initial file
    try:
        h_formula_count = 0
        h_format_count = 0
        for row in range(2, 35):
            cell = ws.cell(row=row, column=8)
            val = cell.value
            nf = cell.number_format

            # Check for severance amount formula: (EX/52)*GX pattern
            if isinstance(val, str):
                val_up = val.upper().replace(' ', '')
                # Allow variations: =(E2/52)*G2 or =E2/52*G2 or similar
                if '/52' in val_up and val_up.startswith('='):
                    h_formula_count += 1

            # Check for currency number format
            if '$#,##0.00' in nf or nf == '$#,##0.00':
                h_format_count += 1

        if h_formula_count == 33 and h_format_count == 33:
            print(f"PASS: Component 3 — All 33 rows have severance amount formula in H with $#,##0.00 format "
                  f"({h_formula_count}/33 formulas, {h_format_count}/33 formatted) (0.25 pts)")
            total_score += 0.25
        elif h_formula_count == 33 and h_format_count >= 25:
            print(f"PARTIAL: Component 3 — Formulas correct ({h_formula_count}/33) but format issues "
                  f"({h_format_count}/33 with $#,##0.00) (0.15 pts)")
            total_score += 0.15
        elif h_formula_count >= 25:
            print(f"PARTIAL: Component 3 — {h_formula_count}/33 rows have amount formula in H (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — Expected amount formulas in H2:H34, found {h_formula_count}/33; "
                  f"currency format: {h_format_count}/33")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Total row — G35 label + H35 SUM formula, both cells bold, H35 currency format (0.25 points)
    # Verifies agent added the totals row — row 35 doesn't exist in initial file
    try:
        g35 = ws.cell(row=35, column=7)
        h35 = ws.cell(row=35, column=8)

        g35_has_label = (isinstance(g35.value, str) and
                         'total' in g35.value.lower() and
                         'severance' in g35.value.lower())
        h35_has_sum = (isinstance(h35.value, str) and
                       h35.value.upper().replace(' ', '').startswith('=SUM(H'))
        h35_bold = (h35.font.bold == True)
        h35_currency = ('$#,##0.00' in h35.number_format)

        components_passed = sum([g35_has_label, h35_has_sum, h35_bold, h35_currency])

        if g35_has_label and h35_has_sum and h35_bold and h35_currency:
            print(f"PASS: Component 4 — Total row complete: G35='{g35.value}', H35='{h35.value}', "
                  f"bold={h35_bold}, format='{h35.number_format}' (0.25 pts)")
            total_score += 0.25
        elif g35_has_label and h35_has_sum:
            print(f"PARTIAL: Component 4 — Label and SUM formula present but formatting issues "
                  f"(bold={h35_bold}, currency={h35_currency}) (0.15 pts)")
            total_score += 0.15
        elif h35_has_sum:
            print(f"PARTIAL: Component 4 — H35 has SUM formula but label missing in G35 "
                  f"(G35='{g35.value}') (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 — Total row incomplete: G35='{g35.value}', H35='{h35.value}', "
                  f"bold={h35_bold}, format='{h35.number_format}'")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
