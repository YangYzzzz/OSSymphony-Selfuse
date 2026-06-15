"""
Reward Script: Fix cross-sheet 3D range references after sheet reorder
Task ID: calc_tbl_095
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Broken 3D range references removed from B2:E13 formulas
  Component 2 (0.4): Corrected formulas reference both Sheet2 and Sheet3 explicitly
  Component 3 (0.2): F column grand-total formulas and row 14 TOTAL formulas preserved
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_095'


def persist_app_state(domain):
    """Save any unsaved LibreOffice state before verification."""
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        import time
        pyautogui.hotkey("ctrl", "s")
        time.sleep(1.0)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sheet1 must exist
    if 'Sheet1' not in wb.sheetnames:
        print("CRITICAL: Sheet1 not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Sheet1']

    # We check the cross-sheet formulas in cells B2:E13 (the summary formulas).
    # Initial (broken): =SUM(Sheet2.B2:Sheet3.B2) — a 3D range reference
    # Golden (fixed): =Sheet2.B2+Sheet3.B2 — explicit per-sheet references
    # Any correct fix eliminates the 3D range pattern and references both sheets.

    cross_sheet_cells = []
    for row in range(2, 14):  # rows 2-13
        for col_letter in ['B', 'C', 'D', 'E']:
            cross_sheet_cells.append(f'{col_letter}{row}')

    total_cross_cells = len(cross_sheet_cells)  # 48 cells

    # Component 1: Broken 3D range references removed (0.4 points)
    # The broken pattern is SUM(Sheet2.XX:Sheet3.XX) or SUM(Sheet3.XX:Sheet2.XX)
    # — any formula using a colon between two sheet references (3D range syntax)
    try:
        no_3d_count = 0
        has_3d_count = 0
        three_d_pattern = re.compile(r'Sheet\d+\.[A-Z]+\d+:Sheet\d+\.[A-Z]+\d+', re.IGNORECASE)

        for coord in cross_sheet_cells:
            val = ws[coord].value
            if val is None:
                continue
            val_str = str(val)
            if three_d_pattern.search(val_str):
                has_3d_count += 1
            else:
                no_3d_count += 1

        if has_3d_count == 0 and no_3d_count > 0:
            print(f"PASS: Component 1 -- All {no_3d_count} cross-sheet formulas have no 3D range refs (0.4 pts)")
            total_score += 0.4
        elif has_3d_count > 0:
            # Partial: proportional to how many were fixed
            fixed_ratio = no_3d_count / total_cross_cells if total_cross_cells > 0 else 0
            partial = round(0.4 * fixed_ratio, 2)
            print(f"FAIL: Component 1 -- {has_3d_count} cells still have 3D range refs, {no_3d_count} fixed. Partial: {partial}")
            total_score += partial
        else:
            print("FAIL: Component 1 -- No formula cells found in cross-sheet range")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Formulas correctly reference both Sheet2 and Sheet3 (0.4 points)
    # Each formula in B2:E13 should reference both Sheet2 and Sheet3 explicitly
    try:
        correct_ref_count = 0
        incorrect_ref_count = 0

        for coord in cross_sheet_cells:
            val = ws[coord].value
            if val is None:
                incorrect_ref_count += 1
                continue
            val_str = str(val).upper()
            # Must be a formula
            if not val_str.startswith('='):
                incorrect_ref_count += 1
                continue
            # Must reference both Sheet2 and Sheet3
            has_sheet2 = 'SHEET2.' in val_str
            has_sheet3 = 'SHEET3.' in val_str
            # Must NOT have the broken 3D range pattern
            has_3d = bool(three_d_pattern.search(str(ws[coord].value)))
            if has_sheet2 and has_sheet3 and not has_3d:
                correct_ref_count += 1
            else:
                incorrect_ref_count += 1

        if correct_ref_count == total_cross_cells:
            print(f"PASS: Component 2 -- All {correct_ref_count} formulas correctly reference Sheet2 and Sheet3 (0.4 pts)")
            total_score += 0.4
        elif correct_ref_count > 0:
            partial = round(0.4 * (correct_ref_count / total_cross_cells), 2)
            print(f"FAIL: Component 2 -- {correct_ref_count}/{total_cross_cells} formulas correct. Partial: {partial}")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- No formulas correctly reference both Sheet2 and Sheet3")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Cross-sheet refs fixed AND structural formulas preserved (0.2 points)
    # This is a compound check: awards points only when BOTH conditions are met:
    #   (a) No 3D range refs remain (cross-sheet fix applied)
    #   (b) F2:F13 and B14:F14 structural formulas are intact
    # This ensures the component only passes on golden (fixed) and fails on initial (broken).
    try:
        # Gate: cross-sheet refs must already be fixed (no 3D ranges)
        if has_3d_count > 0:
            print(f"FAIL: Component 3 -- 3D range refs still present; structural check gated")
        else:
            preserved_count = 0
            total_preservation_checks = 0

            # Check F2:F13 — each should be =SUM(Bn:En)
            for row in range(2, 14):
                total_preservation_checks += 1
                val = ws[f'F{row}'].value
                if val and isinstance(val, str):
                    expected = f'=SUM(B{row}:E{row})'
                    if val.upper().replace(' ', '') == expected.upper().replace(' ', ''):
                        preserved_count += 1

            # Check B14:F14 — TOTAL row
            col_letters = ['B', 'C', 'D', 'E', 'F']
            for cl in col_letters:
                total_preservation_checks += 1
                val = ws[f'{cl}14'].value
                if val and isinstance(val, str):
                    expected = f'=SUM({cl}2:{cl}13)'
                    if val.upper().replace(' ', '') == expected.upper().replace(' ', ''):
                        preserved_count += 1

            if preserved_count == total_preservation_checks:
                print(f"PASS: Component 3 -- Refs fixed AND all {preserved_count} structural formulas preserved (0.2 pts)")
                total_score += 0.2
            elif preserved_count > 0:
                partial = round(0.2 * (preserved_count / total_preservation_checks), 2)
                print(f"FAIL: Component 3 -- {preserved_count}/{total_preservation_checks} structural formulas preserved. Partial: {partial}")
                total_score += partial
            else:
                print(f"FAIL: Component 3 -- No structural formulas (F col, row 14) preserved")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
