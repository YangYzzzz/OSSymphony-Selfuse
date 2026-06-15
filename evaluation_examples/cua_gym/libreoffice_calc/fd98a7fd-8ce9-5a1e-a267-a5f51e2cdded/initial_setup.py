"""
Initial Setup: Cross-sheet 3D reference fix task
Task ID: calc_tbl_095
Domain: libreoffice_calc

Creates a workbook with 3 sheets (Sheet1, Sheet3, Sheet2 — reordered).
Sheet2 has regional sales data, Sheet3 has regional expense data.
Sheet1 (summary) has BROKEN 3D references: =SUM(Sheet2.B2:Sheet3.B2)
which, because sheet order is Sheet1→Sheet3→Sheet2, captures Sheet3 through
Sheet2 (both sheets anyway here, but the 3D range semantics are wrong —
the reference should directly address each sheet, not rely on positional range).
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_095'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Styling helpers ---
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill_blue = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_fill_green = PatternFill(start_color="FF548235", end_color="FF548235", fill_type="solid")
    header_fill_purple = PatternFill(start_color="FF7030A0", end_color="FF7030A0", fill_type="solid")
    data_font = Font(name="Arial", size=11)
    currency_fmt = '#,##0'
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # =========================================================
    # Sheet2: Regional Sales Data
    # =========================================================
    ws2 = wb.active
    ws2.title = "Sheet2"

    sales_headers = ["Region", "Q1 Sales", "Q2 Sales", "Q3 Sales", "Q4 Sales"]
    sales_data = [
        ["North America", 145200, 162800, 158400, 189600],
        ["Europe", 98500, 112300, 107800, 125400],
        ["Asia Pacific", 78900, 85600, 92300, 104500],
        ["Latin America", 42100, 45800, 48200, 56300],
        ["Middle East", 31500, 34200, 37800, 41600],
        ["Africa", 18700, 21300, 23500, 27800],
        ["Eastern Europe", 35600, 38900, 41200, 46700],
        ["South Asia", 52300, 58700, 63400, 71200],
        ["Oceania", 28400, 31200, 33800, 37500],
        ["Central Asia", 15800, 17600, 19200, 22400],
        ["Nordic", 41200, 44800, 47300, 52100],
        ["Southeast Asia", 33600, 37200, 40100, 45800],
    ]

    for c, h in enumerate(sales_headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill_blue
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for r, row_data in enumerate(sales_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = thin_border
            if c >= 2:
                cell.number_format = currency_fmt

    ws2.column_dimensions["A"].width = 18
    for col_letter in ["B", "C", "D", "E"]:
        ws2.column_dimensions[col_letter].width = 14

    # =========================================================
    # Sheet3: Regional Expense Data
    # =========================================================
    ws3 = wb.create_sheet("Sheet3")

    expense_headers = ["Region", "Q1 Expenses", "Q2 Expenses", "Q3 Expenses", "Q4 Expenses"]
    expense_data = [
        ["North America", 89400, 95200, 91800, 102300],
        ["Europe", 62100, 68500, 65200, 74800],
        ["Asia Pacific", 45300, 49800, 53600, 61200],
        ["Latin America", 28700, 31200, 33500, 38100],
        ["Middle East", 19800, 21600, 24100, 26800],
        ["Africa", 12400, 14100, 15800, 18200],
        ["Eastern Europe", 22800, 25100, 27400, 30600],
        ["South Asia", 34500, 38200, 41800, 46300],
        ["Oceania", 17600, 19800, 21500, 24100],
        ["Central Asia", 10200, 11800, 13100, 15200],
        ["Nordic", 26800, 29400, 31200, 34700],
        ["Southeast Asia", 21400, 24100, 26800, 30200],
    ]

    for c, h in enumerate(expense_headers, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill_green
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for r, row_data in enumerate(expense_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = thin_border
            if c >= 2:
                cell.number_format = currency_fmt

    ws3.column_dimensions["A"].width = 18
    for col_letter in ["B", "C", "D", "E"]:
        ws3.column_dimensions[col_letter].width = 14

    # =========================================================
    # Sheet1: Summary with BROKEN 3D references
    # =========================================================
    ws1 = wb.create_sheet("Sheet1")

    summary_headers = ["Region", "Total Q1", "Total Q2", "Total Q3", "Total Q4", "Grand Total"]
    for c, h in enumerate(summary_headers, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill_purple
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Region names in column A
    regions = [row[0] for row in sales_data]
    for r, region in enumerate(regions, 2):
        cell = ws1.cell(row=r, column=1, value=region)
        cell.font = data_font
        cell.border = thin_border

    # BROKEN 3D references: =SUM(Sheet2.B2:Sheet3.B2)
    # In LibreOffice, 3D references use Sheet2.Cell:Sheet3.Cell syntax
    # meaning "sum across all sheets from Sheet2 to Sheet3 in positional order"
    # With sheet order Sheet1, Sheet3, Sheet2 — this range goes Sheet2→Sheet2
    # (or Sheet3→Sheet2 depending on direction), which is WRONG.
    # The correct approach would be explicit per-sheet references.
    col_letters = ["B", "C", "D", "E"]
    for r in range(2, len(regions) + 2):
        for ci, col in enumerate(col_letters):
            cell_ref = f"{col}{r}"
            # Broken 3D reference
            formula = f"=SUM(Sheet2.{cell_ref}:Sheet3.{cell_ref})"
            cell = ws1.cell(row=r, column=ci + 2, value=formula)
            cell.font = data_font
            cell.border = thin_border
            cell.number_format = currency_fmt

        # Grand Total = sum of Q1-Q4 for this row
        grand_cell = ws1.cell(row=r, column=6, value=f"=SUM(B{r}:E{r})")
        grand_cell.font = Font(name="Arial", size=11, bold=True)
        grand_cell.border = thin_border
        grand_cell.number_format = currency_fmt

    # Total row at the bottom
    total_row = len(regions) + 2
    ws1.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", size=11, bold=True)
    ws1.cell(row=total_row, column=1).border = thin_border
    for c in range(2, 7):
        col_letter = chr(64 + c)
        cell = ws1.cell(row=total_row, column=c,
                        value=f"=SUM({col_letter}2:{col_letter}{total_row - 1})")
        cell.font = Font(name="Arial", size=11, bold=True)
        cell.border = thin_border
        cell.number_format = currency_fmt

    ws1.column_dimensions["A"].width = 18
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws1.column_dimensions[col_letter].width = 14

    # =========================================================
    # Reorder sheets: Sheet1, Sheet3, Sheet2
    # =========================================================
    # Current order after creation: Sheet2 (index 0), Sheet3 (index 1), Sheet1 (index 2)
    # Target: Sheet1 (index 0), Sheet3 (index 1), Sheet2 (index 2)
    wb.move_sheet("Sheet1", offset=-2)  # Sheet1 to front → Sheet1, Sheet2, Sheet3
    wb.move_sheet("Sheet3", offset=-1)  # Swap Sheet3 before Sheet2 → Sheet1, Sheet3, Sheet2

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet order: {wb.sheetnames}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
