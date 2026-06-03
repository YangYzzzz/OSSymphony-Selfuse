"""
Initial Setup: HR Certification Expiry Tracker
Task ID: calc_hr_certification_expiry_024
Domain: libreoffice_calc

Creates a Certifications sheet with employee certification records.
Column F (Days Until Expiry) is empty - no formulas, no conditional formatting.
"""

import openpyxl
from datetime import date, timedelta
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_certification_expiry_024'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Certifications'

    # Headers
    headers = ['Emp ID', 'Name', 'Certification', 'Issue Date', 'Expiry Date', 'Days Until Expiry']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic employee data - 75 records (rows 2-76)
    first_names = [
        'Sarah', 'Marcus', 'Jennifer', 'David', 'Emily', 'Robert', 'Linda', 'James',
        'Patricia', 'Michael', 'Barbara', 'William', 'Susan', 'Richard', 'Jessica',
        'Thomas', 'Karen', 'Charles', 'Lisa', 'Daniel', 'Nancy', 'Matthew', 'Betty',
        'Anthony', 'Margaret', 'Donald', 'Sandra', 'Mark', 'Ashley', 'Paul',
        'Kimberly', 'Steven', 'Donna', 'Andrew', 'Carol', 'Kenneth', 'Michelle',
        'Joshua', 'Amanda', 'Kevin', 'Dorothy', 'Brian', 'Melissa', 'George',
        'Deborah', 'Timothy', 'Stephanie', 'Ronald', 'Rebecca', 'Edward'
    ]
    last_names = [
        'Chen', 'Johnson', 'Williams', 'Brown', 'Jones', 'Garcia', 'Miller', 'Davis',
        'Wilson', 'Martinez', 'Anderson', 'Taylor', 'Thomas', 'Hernandez', 'Moore',
        'Jackson', 'Thompson', 'White', 'Lopez', 'Lee', 'Harris', 'Clark', 'Lewis',
        'Robinson', 'Walker', 'Hall', 'Allen', 'Young', 'King', 'Wright',
        'Scott', 'Torres', 'Nguyen', 'Hill', 'Flores', 'Green', 'Adams', 'Nelson',
        'Baker', 'Carter', 'Rivera', 'Mitchell', 'Perez', 'Roberts', 'Turner'
    ]
    certifications = [
        'PMP', 'AWS Certified Solutions Architect', 'CISSP', 'Six Sigma Green Belt',
        'Six Sigma Black Belt', 'CPA', 'CFA', 'PHR', 'SPHR', 'SHRM-CP',
        'CompTIA Security+', 'CompTIA Network+', 'Certified Scrum Master',
        'ITIL Foundation', 'ITIL Practitioner', 'Oracle DBA Certified',
        'Microsoft Azure Administrator', 'Google Cloud Professional',
        'Salesforce Administrator', 'ISO 9001 Lead Auditor',
        'OSHA 30-Hour General Industry', 'OSHA 10-Hour Construction',
        'First Aid/CPR Certified', 'Lean Practitioner', 'Data Science Professional',
        'Kubernetes Administrator', 'Docker Certified Associate',
        'Certified Ethical Hacker', 'CISM', 'CISA'
    ]

    # Today's reference date
    today = date(2026, 3, 4)

    # Generate varied expiry dates:
    # ~10 already expired, ~10 expiring within 30 days, ~15 expiring within 90 days, ~40 valid
    expiry_profiles = (
        [-45, -30, -20, -15, -10, -7, -5, -3, -2, -1] +           # 10 expired
        [5, 10, 15, 18, 20, 22, 25, 28, 29, 30] +                   # 10 expiring <=30 days
        [35, 40, 45, 50, 55, 60, 65, 70, 75, 80, 85, 88, 89, 90] +  # 14 expiring 31-90
        [95, 100, 120, 150, 180, 200, 240, 270, 300, 330,            # 41 valid >90
         365, 400, 450, 500, 550, 600, 730, 365, 400, 420,
         180, 200, 210, 240, 365, 400, 450, 500, 550, 600, 180,
         91, 110, 130, 160, 190, 220, 260, 290, 310, 340]
    )

    random.seed(42)
    used_names = []

    for i in range(75):
        emp_id = f'EMP{1000 + i + 1}'
        fn = first_names[i % len(first_names)]
        ln = last_names[i % len(last_names)]
        name = f'{fn} {ln}'
        cert = certifications[i % len(certifications)]

        days_offset = expiry_profiles[i]
        expiry_date = today + timedelta(days=days_offset)
        # Issue date: certification valid for 3 years, so issue = expiry - 3 years
        issue_date = expiry_date - timedelta(days=3*365)

        row = i + 2
        ws.cell(row=row, column=1, value=emp_id)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=cert)
        ws.cell(row=row, column=4, value=issue_date)
        ws.cell(row=row, column=5, value=expiry_date)
        # Column F intentionally left empty

    # Set date format for columns D and E
    from openpyxl.styles import numbers
    for row in range(2, 77):
        ws.cell(row=row, column=4).number_format = 'yyyy-mm-dd'
        ws.cell(row=row, column=5).number_format = 'yyyy-mm-dd'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 32
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheets: {wb.sheetnames}')
    print(f'Rows: {ws.max_row}')
    print(f'Columns: {ws.max_column}')
    print(f'Column F (Days Until Expiry) is empty - no formulas added')

create_initial()
