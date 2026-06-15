"""
Reward Script: HR Certification Expiry Tracker
Task ID: calc_hr_certification_expiry_024
Domain: libreoffice_calc
Scoring:
  - Component 1: F2:F76 contains =En-TODAY() formulas with integer number format (0.35 pts)
  - Component 2: CF rule for expired (<0): bg #FF0000, font #FFFFFF (0.20 pts)
  - Component 3: CF rule for expiring soon (0-30 days): bg #FF6600 orange (0.20 pts)
  - Component 4: CF rule for expiring within 90 days (31-90): bg #FFFF00 yellow (0.15 pts)
  - Component 5: CF rule for valid (>90 days): bg #70AD47 green (0.10 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_certification_expiry_024'


def normalize_formula(f):
    """Normalize formula string for comparison: uppercase, no spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Certifications' sheet must exist
    if 'Certifications' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Certifications' not found.")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Certifications']

    # Component 1: F2:F76 must contain =En-TODAY() formulas with integer ('0') number format (0.35 points)
    # The initial file has column F completely empty — so any formula there represents the task change.
    try:
        formula_correct_count = 0
        format_correct_count = 0
        total_rows = 75  # rows 2–76

        for row in range(2, 77):
            cell = ws.cell(row=row, column=6)
            val = cell.value
            expected_formula = f'=E{row}-TODAY()'

            # Check formula (case-insensitive, whitespace-insensitive)
            if normalize_formula(val) == normalize_formula(expected_formula):
                formula_correct_count += 1

            # Check number format is integer '0'
            if cell.number_format in ('0', '0.0', '#,##0', 'General') and cell.number_format == '0':
                format_correct_count += 1

        formula_ratio = formula_correct_count / total_rows
        format_ratio = format_correct_count / total_rows

        if formula_ratio >= 1.0:
            print(f"PASS: Component 1a — All F2:F76 contain correct =En-TODAY() formula ({formula_correct_count}/{total_rows})")
            formula_pts = 0.25
        elif formula_ratio >= 0.5:
            print(f"PARTIAL: Component 1a — {formula_correct_count}/{total_rows} formulas correct")
            formula_pts = 0.10
        else:
            print(f"FAIL: Component 1a — Only {formula_correct_count}/{total_rows} formulas correct")
            formula_pts = 0.0

        if format_ratio >= 1.0:
            print(f"PASS: Component 1b — All F2:F76 formatted as integer '0' ({format_correct_count}/{total_rows})")
            format_pts = 0.10
        else:
            print(f"FAIL: Component 1b — Only {format_correct_count}/{total_rows} cells have number_format='0'")
            format_pts = 0.0

        total_score += formula_pts + format_pts

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: CF rule — expired (F<0): background #FF0000, font #FFFFFF (0.20 points)
    # The initial file has 0 CF rules, so any CF rule is a task-introduced change.
    try:
        cf_rules_dict = ws.conditional_formatting._cf_rules
        all_rules = []
        for cf_range, rules in cf_rules_dict.items():
            for rule in rules:
                all_rules.append((str(cf_range), rule))

        # Look for the expired rule: formula '$F2<0', fill FFFF0000, font FFFFFFFF
        found_expired_rule = False
        for cf_range_str, rule in all_rules:
            if rule.type == 'expression' and rule.formula:
                formula_str = normalize_formula(rule.formula[0])
                if formula_str == normalize_formula('$F2<0'):
                    # Check fill color
                    fill_ok = False
                    font_ok = False
                    if rule.dxf and rule.dxf.fill:
                        fg = rule.dxf.fill.fgColor.rgb
                        if fg in ('FFFF0000', 'FF0000'):
                            fill_ok = True
                    if rule.dxf and rule.dxf.font and rule.dxf.font.color:
                        fc = rule.dxf.font.color.rgb
                        if fc in ('FFFFFFFF', 'FFFFFF'):
                            font_ok = True

                    if fill_ok and font_ok:
                        found_expired_rule = True
                        print(f"PASS: Component 2 — Expired rule ($F2<0): bg=#FF0000, font=#FFFFFF found")
                    elif fill_ok:
                        found_expired_rule = True
                        print(f"PARTIAL: Component 2 — Expired rule ($F2<0): bg=#FF0000 found but font color missing/wrong")
                    else:
                        print(f"FAIL: Component 2 — Expired rule ($F2<0) found but fill={rule.dxf.fill.fgColor.rgb if rule.dxf and rule.dxf.fill else None}")

        if found_expired_rule:
            total_score += 0.20
        else:
            print("FAIL: Component 2 — No CF rule with formula '$F2<0' found")

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: CF rule — expiring soon (AND($F2>=0,$F2<=30)): background #FF6600 orange (0.20 points)
    try:
        found_orange_rule = False
        for cf_range_str, rule in all_rules:
            if rule.type == 'expression' and rule.formula:
                formula_str = normalize_formula(rule.formula[0])
                # Match AND($F2>=0,$F2<=30) in various forms
                if 'F2>=0' in formula_str and 'F2<=30' in formula_str:
                    fill_ok = False
                    if rule.dxf and rule.dxf.fill:
                        fg = rule.dxf.fill.fgColor.rgb
                        if fg in ('FFFF6600', 'FF6600'):
                            fill_ok = True
                    if fill_ok:
                        found_orange_rule = True
                        print(f"PASS: Component 3 — Orange rule (0-30 days): bg=#FF6600 found")
                    else:
                        print(f"FAIL: Component 3 — Rule for 0-30 days found but fill color wrong: {rule.dxf.fill.fgColor.rgb if rule.dxf and rule.dxf.fill else None}")

        if found_orange_rule:
            total_score += 0.20
        else:
            print("FAIL: Component 3 — No CF rule for 0-30 days (AND($F2>=0,$F2<=30)) found")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: CF rule — expiring within 90 days (AND($F2>30,$F2<=90)): background #FFFF00 yellow (0.15 points)
    try:
        found_yellow_rule = False
        for cf_range_str, rule in all_rules:
            if rule.type == 'expression' and rule.formula:
                formula_str = normalize_formula(rule.formula[0])
                if 'F2>30' in formula_str and 'F2<=90' in formula_str:
                    fill_ok = False
                    if rule.dxf and rule.dxf.fill:
                        fg = rule.dxf.fill.fgColor.rgb
                        if fg in ('FFFFFF00', 'FFFF00'):
                            fill_ok = True
                    if fill_ok:
                        found_yellow_rule = True
                        print(f"PASS: Component 4 — Yellow rule (31-90 days): bg=#FFFF00 found")
                    else:
                        print(f"FAIL: Component 4 — Rule for 31-90 days found but fill color wrong: {rule.dxf.fill.fgColor.rgb if rule.dxf and rule.dxf.fill else None}")

        if found_yellow_rule:
            total_score += 0.15
        else:
            print("FAIL: Component 4 — No CF rule for 31-90 days (AND($F2>30,$F2<=90)) found")

    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: CF rule — valid (F>90): background #70AD47 green (0.10 points)
    try:
        found_green_rule = False
        for cf_range_str, rule in all_rules:
            if rule.type == 'expression' and rule.formula:
                formula_str = normalize_formula(rule.formula[0])
                if normalize_formula('$F2>90') == formula_str or formula_str == normalize_formula('F2>90'):
                    fill_ok = False
                    if rule.dxf and rule.dxf.fill:
                        fg = rule.dxf.fill.fgColor.rgb
                        if fg in ('FF70AD47', '70AD47'):
                            fill_ok = True
                    if fill_ok:
                        found_green_rule = True
                        print(f"PASS: Component 5 — Green rule (>90 days): bg=#70AD47 found")
                    else:
                        print(f"FAIL: Component 5 — Rule for >90 days found but fill color wrong: {rule.dxf.fill.fgColor.rgb if rule.dxf and rule.dxf.fill else None}")

        if found_green_rule:
            total_score += 0.10
        else:
            print("FAIL: Component 5 — No CF rule for >90 days ($F2>90) found")

    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
