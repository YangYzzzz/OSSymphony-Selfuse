"""
Reward Script: Adjust column widths to optimal width for all columns
Task ID: calc_gfl_092
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): Previously too-narrow columns (A,D,H,I) are now wider
  Component 2 (0.35): Previously too-wide columns (E,G) are now narrower
  Component 3 (0.30): All columns have width >= max content length (no truncation)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_092'


def get_max_content_lengths(ws):
    """Compute the max string length of content in each column."""
    max_lens = {}
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                l = len(str(val))
                if l > max_len:
                    max_len = l
        col_letter = chr(64 + col_idx)
        max_lens[col_letter] = max_len
    return max_lens


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Inventory' sheet must exist
    if 'Inventory' not in wb.sheetnames:
        print("FAIL: 'Inventory' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Inventory']

    # Precondition: data integrity - must have 46 rows and 10 columns
    if ws.max_row < 45 or ws.max_column < 10:
        print(f"FAIL: Data integrity - expected >=45 rows and >=10 cols, found {ws.max_row} rows, {ws.max_column} cols")
        print("REWARD: 0.0")
        return 0.0

    max_lens = get_max_content_lengths(ws)

    # Known initial widths (from setup-gen's initial file)
    initial_widths = {
        'A': 6.0, 'B': 8.0, 'C': 10.0, 'D': 6.0, 'E': 30.0,
        'F': 8.0, 'G': 25.0, 'H': 4.0, 'I': 5.0, 'J': 8.0
    }

    # Get current widths
    current_widths = {}
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        dim = ws.column_dimensions.get(col_letter)
        if dim and dim.width is not None:
            current_widths[col_letter] = dim.width
        else:
            current_widths[col_letter] = 8.43  # default Excel width
        print(f"  Col {col_letter}: width={current_widths[col_letter]}, max_content_len={max_lens.get(col_letter, 0)}")

    print()

    # Component 1: Previously too-narrow columns are now wider (0.35 points)
    # Columns A(6->13), B(8->30), D(6->14), H(4->7), I(5->16) were too narrow
    # for their content. After optimal width, they should be significantly wider.
    try:
        narrow_cols = ['A', 'B', 'D', 'H', 'I']
        widened_count = 0
        for col in narrow_cols:
            init_w = initial_widths[col]
            curr_w = current_widths[col]
            content_len = max_lens[col]
            # The column should now be wider than initial AND at least as wide as content length
            if curr_w > init_w and curr_w >= content_len:
                widened_count += 1
                print(f"PASS: Col {col} widened from {init_w} to {curr_w} (content needs ~{content_len})")
            else:
                print(f"FAIL: Col {col} not properly widened: init={init_w}, current={curr_w}, content_len={content_len}")

        if widened_count == len(narrow_cols):
            print(f"PASS: Component 1 - All {len(narrow_cols)} narrow columns widened (0.35 pts)")
            total_score += 0.35
        elif widened_count >= 3:
            partial = round(0.35 * widened_count / len(narrow_cols), 2)
            print(f"PARTIAL: Component 1 - {widened_count}/{len(narrow_cols)} narrow columns widened ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - Only {widened_count}/{len(narrow_cols)} narrow columns widened")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Previously too-wide columns are now narrower (0.35 points)
    # Columns E(30->21) and G(25->11) had excessive width compared to content.
    # After optimal width, they should be closer to content length.
    try:
        wide_cols = {'E': 30.0, 'G': 25.0}
        narrowed_count = 0
        for col, init_w in wide_cols.items():
            curr_w = current_widths[col]
            content_len = max_lens[col]
            # Column should now be narrower than before AND still accommodate content
            if curr_w < init_w and curr_w >= content_len:
                narrowed_count += 1
                print(f"PASS: Col {col} narrowed from {init_w} to {curr_w} (content needs ~{content_len})")
            else:
                print(f"FAIL: Col {col} not properly narrowed: init={init_w}, current={curr_w}, content_len={content_len}")

        if narrowed_count == len(wide_cols):
            print(f"PASS: Component 2 - All {len(wide_cols)} wide columns narrowed (0.35 pts)")
            total_score += 0.35
        elif narrowed_count >= 1:
            partial = round(0.35 * narrowed_count / len(wide_cols), 2)
            print(f"PARTIAL: Component 2 - {narrowed_count}/{len(wide_cols)} wide columns narrowed ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 - No wide columns were narrowed")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: All columns have width >= max content length (0.30 points)
    # This ensures no content is truncated after the optimization.
    # On initial_env, several columns are too narrow for their content, so this FAILS.
    try:
        all_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
        fit_count = 0
        for col in all_cols:
            curr_w = current_widths[col]
            content_len = max_lens[col]
            if curr_w >= content_len:
                fit_count += 1
            else:
                print(f"FAIL: Col {col} width {curr_w} < content len {content_len} (truncated)")

        if fit_count == len(all_cols):
            print(f"PASS: Component 3 - All {len(all_cols)} columns fit their content (0.30 pts)")
            total_score += 0.30
        elif fit_count >= 7:
            partial = round(0.30 * fit_count / len(all_cols), 2)
            print(f"PARTIAL: Component 3 - {fit_count}/{len(all_cols)} columns fit content ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 - Only {fit_count}/{len(all_cols)} columns fit their content")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
