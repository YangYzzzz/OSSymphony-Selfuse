"""
Initial Setup: Create inventory spreadsheet with non-optimal column widths
Task ID: calc_gfl_092
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_092'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventory'

    # Headers
    headers = [
        'Product ID', 'Name', 'Description', 'SKU', 'Category',
        'Supplier', 'Unit Cost', 'Stock', 'Reorder Point', 'Location'
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Realistic inventory data - 45 rows
    data = [
        ['PRD-001', 'Industrial Bearing Assembly', 'High-precision steel bearing assembly for CNC machines, rated for 10000 RPM', 'SKU-BRG-4421', 'Machine Parts', 'Precision Components Inc.', 145.99, 230, 50, 'Warehouse A - Shelf 12'],
        ['PRD-002', 'LED Panel Light 60W', 'Commercial grade LED panel, 4000K neutral white, 6000 lumens', 'SKU-LED-7823', 'Lighting', 'BrightTech Solutions', 89.50, 485, 100, 'Warehouse B - Bay 3'],
        ['PRD-003', 'Hydraulic Cylinder Rod', 'Chrome-plated rod, 2-inch bore, 18-inch stroke length', 'SKU-HYD-1190', 'Hydraulics', 'FluidPower Systems LLC', 312.75, 42, 15, 'Warehouse A - Shelf 8'],
        ['PRD-004', 'Stainless Steel Pipe Fitting', '316L stainless steel elbow fitting, 1.5 inch NPT threading', 'SKU-PIP-3345', 'Plumbing', 'MetalWorks Supply Co.', 28.40, 1250, 200, 'Warehouse C - Rack 5'],
        ['PRD-005', 'Thermal Insulation Roll', 'Fiberglass insulation, R-30, 15-inch width, 40 sq ft roll', 'SKU-INS-5567', 'Insulation', 'ThermoGuard Industries', 67.25, 310, 75, 'Warehouse B - Bay 7'],
        ['PRD-006', 'Circuit Breaker 30A', 'Single-pole circuit breaker, 120/240V, bolt-on type', 'SKU-ELC-8890', 'Electrical', 'PowerSafe Electric', 18.75, 890, 150, 'Warehouse A - Shelf 3'],
        ['PRD-007', 'Welding Wire MIG 0.035', 'ER70S-6 mild steel welding wire, 33 lb spool', 'SKU-WLD-2234', 'Welding Supplies', 'ArcMaster Welding', 52.30, 165, 40, 'Warehouse C - Rack 2'],
        ['PRD-008', 'Safety Goggles Anti-Fog', 'ANSI Z87.1 rated, indirect ventilation, polycarbonate lens', 'SKU-SAF-6678', 'Safety Equipment', 'SafeView Protective Gear', 12.95, 2100, 500, 'Warehouse B - Bay 1'],
        ['PRD-009', 'Pneumatic Air Hose 50ft', '3/8 inch ID, 300 PSI rated, with brass fittings', 'SKU-PNU-4456', 'Pneumatics', 'AirFlow Dynamics', 34.80, 275, 60, 'Warehouse A - Shelf 15'],
        ['PRD-010', 'Diamond Cutting Blade 14in', 'Segmented rim, for concrete and masonry, wet/dry cutting', 'SKU-CUT-9912', 'Cutting Tools', 'DiamondEdge Tools', 189.99, 58, 20, 'Warehouse C - Rack 9'],
        ['PRD-011', 'Epoxy Adhesive 2-Part', 'Industrial strength, 24-hour cure time, 3500 PSI tensile strength', 'SKU-ADH-3378', 'Adhesives', 'BondStrong Chemical', 24.50, 620, 100, 'Warehouse B - Bay 4'],
        ['PRD-012', 'Conveyor Belt Roller', 'Galvanized steel, 20-inch face width, 1000 lb capacity per roller', 'SKU-CNV-7745', 'Conveyor Parts', 'ConveyTech Manufacturing', 78.60, 140, 30, 'Warehouse A - Shelf 6'],
        ['PRD-013', 'PLC Controller Module', 'Programmable logic controller, 16 digital inputs, 8 relay outputs', 'SKU-PLC-1123', 'Automation', 'AutoLogic Controls', 425.00, 35, 10, 'Warehouse C - Rack 1'],
        ['PRD-014', 'Rubber Gasket Sheet', 'Neoprene rubber, 1/8 inch thick, 36x36 inch sheet, oil resistant', 'SKU-GSK-5590', 'Sealing', 'SealPro Rubber Works', 42.15, 380, 80, 'Warehouse B - Bay 6'],
        ['PRD-015', 'Torque Wrench 1/2 Drive', 'Click-type, 10-150 ft-lbs range, calibration certificate included', 'SKU-TRQ-8834', 'Hand Tools', 'TorqueMaster Tools', 95.40, 92, 25, 'Warehouse A - Shelf 10'],
        ['PRD-016', 'Cable Tray 12in Wide', 'Galvanized steel ladder type, 10-foot section, NEMA rated', 'SKU-CBL-2267', 'Cable Management', 'WireWay Solutions', 56.80, 210, 50, 'Warehouse C - Rack 4'],
        ['PRD-017', 'Pressure Gauge 0-300PSI', 'Bourdon tube type, 4-inch dial, 1/4 NPT connection', 'SKU-PRG-6641', 'Instrumentation', 'MeasureTech Instruments', 38.25, 445, 75, 'Warehouse B - Bay 2'],
        ['PRD-018', 'Forklift Battery 36V', 'Deep cycle lead-acid, 750 Ah capacity, with watering system', 'SKU-BAT-9908', 'Batteries', 'PowerCell Energy', 2850.00, 12, 5, 'Warehouse A - Shelf 1'],
        ['PRD-019', 'Anti-Vibration Mount', 'Rubber-to-metal bonded, 500 lb capacity, M12 stud', 'SKU-VIB-4423', 'Vibration Control', 'DampTech Engineering', 31.60, 520, 100, 'Warehouse C - Rack 7'],
        ['PRD-020', 'Fire Extinguisher 10lb', 'ABC dry chemical, wall mount bracket included, DOT approved', 'SKU-FIR-7756', 'Fire Safety', 'FireGuard Safety', 65.90, 180, 40, 'Warehouse B - Bay 8'],
        ['PRD-021', 'Stainless Steel Sheet 4x8', '304 grade, 16 gauge, #4 brushed finish, mill certified', 'SKU-STL-3312', 'Sheet Metal', 'MetalWorks Supply Co.', 385.00, 28, 10, 'Warehouse A - Shelf 2'],
        ['PRD-022', 'Air Compressor Filter', 'Coalescing element, 0.01 micron, removes oil and moisture', 'SKU-FLT-5589', 'Filtration', 'AirFlow Dynamics', 47.30, 340, 60, 'Warehouse C - Rack 3'],
        ['PRD-023', 'Industrial Caster 6in', 'Polyurethane wheel, swivel with brake, 1200 lb capacity', 'SKU-CST-8867', 'Material Handling', 'RollEasy Industrial', 29.75, 680, 120, 'Warehouse B - Bay 5'],
        ['PRD-024', 'Thermocouple Type K', 'Stainless steel sheath, 12-inch probe, -200C to 1250C range', 'SKU-TMP-1145', 'Instrumentation', 'MeasureTech Instruments', 55.80, 195, 40, 'Warehouse A - Shelf 14'],
        ['PRD-025', 'Chain Hoist 2-Ton', 'Manual chain block, 10-foot lift height, overload protection', 'SKU-HST-6623', 'Lifting Equipment', 'LiftMaster Heavy Duty', 278.50, 24, 8, 'Warehouse C - Rack 8'],
        ['PRD-026', 'PVC Conduit 3/4in', 'Schedule 40, 10-foot stick, UL listed for electrical use', 'SKU-PVC-4401', 'Electrical', 'PowerSafe Electric', 4.25, 3200, 500, 'Warehouse B - Bay 9'],
        ['PRD-027', 'Drill Bit Set HSS', '29-piece set, 1/16 to 1/2 inch, 135-degree split point', 'SKU-DRL-9978', 'Cutting Tools', 'DiamondEdge Tools', 68.90, 145, 30, 'Warehouse A - Shelf 11'],
        ['PRD-028', 'Hydraulic Hose Assembly', '3/8 inch ID, 4000 PSI, 6-foot length with JIC fittings', 'SKU-HOS-2256', 'Hydraulics', 'FluidPower Systems LLC', 58.40, 290, 50, 'Warehouse C - Rack 6'],
        ['PRD-029', 'Lockout Tagout Kit', '45-piece kit with padlocks, hasps, tags, and carrying case', 'SKU-LOK-7734', 'Safety Equipment', 'SafeView Protective Gear', 124.50, 75, 20, 'Warehouse B - Bay 1'],
        ['PRD-030', 'Gear Motor 1HP', '1725 RPM, 56C frame, TEFC enclosure, 208-230/460V', 'SKU-MTR-3390', 'Motors', 'MotorDrive Technologies', 398.00, 18, 6, 'Warehouse A - Shelf 5'],
        ['PRD-031', 'Sandblasting Media 50lb', 'Aluminum oxide, 80 grit, reusable abrasive media', 'SKU-SND-5512', 'Surface Preparation', 'BlastClean Abrasives', 35.60, 410, 80, 'Warehouse C - Rack 10'],
        ['PRD-032', 'Proximity Sensor Inductive', 'M18 barrel, 8mm sensing distance, PNP NO, IP67 rated', 'SKU-SNS-8845', 'Automation', 'AutoLogic Controls', 42.90, 260, 50, 'Warehouse B - Bay 3'],
        ['PRD-033', 'Steel Shelving Unit 48x24', '5-tier, 4000 lb total capacity, boltless assembly', 'SKU-SHV-1167', 'Storage', 'RackMaster Storage', 189.99, 55, 15, 'Warehouse A - Shelf 7'],
        ['PRD-034', 'Compressed Air Regulator', '1/2 inch NPT, 0-125 PSI output, built-in gauge', 'SKU-REG-6634', 'Pneumatics', 'AirFlow Dynamics', 48.75, 305, 60, 'Warehouse C - Rack 2'],
        ['PRD-035', 'Pallet Jack Standard', '5500 lb capacity, 48-inch fork length, nylon wheels', 'SKU-PLT-9901', 'Material Handling', 'RollEasy Industrial', 425.00, 8, 3, 'Warehouse B - Bay 10'],
        ['PRD-036', 'Wire Terminal Assortment', '1200-piece kit, ring, spade, butt connectors, 22-10 AWG', 'SKU-WIR-2278', 'Electrical', 'PowerSafe Electric', 32.50, 490, 100, 'Warehouse A - Shelf 4'],
        ['PRD-037', 'Lubricating Grease Cartridge', 'NLGI #2, lithium complex, extreme pressure, 14 oz tube', 'SKU-LUB-4489', 'Lubricants', 'LubeMax Industrial', 8.95, 1800, 300, 'Warehouse C - Rack 5'],
        ['PRD-038', 'Emergency Stop Button', 'Mushroom head, twist-to-release, 1 NC + 1 NO contacts', 'SKU-EST-7712', 'Safety Equipment', 'SafeView Protective Gear', 22.40, 350, 60, 'Warehouse B - Bay 2'],
        ['PRD-039', 'V-Belt A68', 'Classical A-section, 68-inch outside length, aramid cord', 'SKU-BLT-3356', 'Power Transmission', 'BeltDrive Components', 14.80, 720, 150, 'Warehouse A - Shelf 13'],
        ['PRD-040', 'Digital Multimeter', 'True RMS, auto-ranging, CAT III 1000V rated, with case', 'SKU-DMM-5534', 'Test Equipment', 'MeasureTech Instruments', 156.00, 68, 15, 'Warehouse C - Rack 1'],
        ['PRD-041', 'Spray Paint Industrial', 'Rust preventative enamel, safety yellow, 12 oz aerosol can', 'SKU-SPR-8801', 'Coatings', 'CoatMaster Paint', 7.50, 2400, 400, 'Warehouse B - Bay 6'],
        ['PRD-042', 'Linear Bearing Rail 500mm', 'MGN12 profile, with carriage block, C3 preload', 'SKU-LBR-1134', 'Machine Parts', 'Precision Components Inc.', 38.90, 185, 35, 'Warehouse A - Shelf 9'],
        ['PRD-043', 'Teflon Tape 1/2in', 'PTFE thread seal tape, 520-inch roll, 3.5 mil thickness', 'SKU-TFN-6690', 'Plumbing', 'SealPro Rubber Works', 3.25, 4500, 800, 'Warehouse C - Rack 4'],
        ['PRD-044', 'Step Ladder 6ft', 'Fiberglass, Type IA 300 lb rating, tool tray top', 'SKU-LDR-9945', 'Access Equipment', 'ClimbSafe Ladders', 145.00, 42, 10, 'Warehouse B - Bay 7'],
        ['PRD-045', 'Hydraulic Jack 20-Ton', 'Bottle type, 7.5-inch lift range, safety overload valve', 'SKU-JCK-2289', 'Lifting Equipment', 'LiftMaster Heavy Duty', 198.50, 30, 8, 'Warehouse A - Shelf 1'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set deliberately BAD column widths - some too narrow, some too wide
    # This is the key: the task asks the user to fix these to optimal width
    ws.column_dimensions['A'].width = 6    # Too narrow for "Product ID" / "PRD-001"
    ws.column_dimensions['B'].width = 8    # Too narrow for long product names
    ws.column_dimensions['C'].width = 10   # Way too narrow for descriptions
    ws.column_dimensions['D'].width = 6    # Too narrow for SKU codes
    ws.column_dimensions['E'].width = 30   # Too wide for category names
    ws.column_dimensions['F'].width = 8    # Too narrow for supplier names
    ws.column_dimensions['G'].width = 25   # Too wide for unit cost numbers
    ws.column_dimensions['H'].width = 4    # Too narrow for stock numbers
    ws.column_dimensions['I'].width = 5    # Too narrow for reorder point
    ws.column_dimensions['J'].width = 8    # Too narrow for location strings

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
