"""
Reward Script: Create custom fill series FQ1-FQ4 and fill A2:A13 with 3 cycles
Task ID: calc_dop_fillseries_custom_052
Domain: libreoffice_calc
Scoring:
  - Component 1: First cycle A2:A5 = FQ1, FQ2, FQ3, FQ4         (0.30 pts)
  - Component 2: Second cycle A6:A9 = FQ1, FQ2, FQ3, FQ4        (0.35 pts)
  - Component 3: Third cycle A10:A13 = FQ1, FQ2, FQ3, FQ4       (0.35 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_dop_fillseries_custom_052'
SHEET_NAME = 'FiscalData'

# Expected fiscal quarter sequence (3 full cycles of FQ1-FQ4)
EXPECTED_CYCLE = ['FQ1', 'FQ2', 'FQ3', 'FQ4']


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.

    Task: Fill A2:A13 in sheet 'FiscalData' with 3 complete cycles of FQ1-FQ4.
    Expected values:
      A2:FQ1, A3:FQ2, A4:FQ3, A5:FQ4,
      A6:FQ1, A7:FQ2, A8:FQ3, A9:FQ4,
      A10:FQ1, A11:FQ2, A12:FQ3, A13:FQ4

    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook (precondition gate)
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists (precondition gate)
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Component 1: First cycle A2:A5 = FQ1, FQ2, FQ3, FQ4 (0.30 points)
    # This FAILS on initial (all None) → PASSES on golden (FQ1-FQ4 in A2:A5)
    try:
        first_cycle_values = [ws.cell(row=r, column=1).value for r in range(2, 6)]
        expected_first = EXPECTED_CYCLE  # ['FQ1', 'FQ2', 'FQ3', 'FQ4']

        if first_cycle_values == expected_first:
            print(f"PASS: Component 1 — First cycle A2:A5 = {first_cycle_values} (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 — Expected A2:A5 = {expected_first}, found {first_cycle_values}")
    except Exception as e:
        print(f"ERROR: Component 1 — Cannot read A2:A5: {e}")

    # Component 2: Second cycle A6:A9 = FQ1, FQ2, FQ3, FQ4 (0.35 points)
    # This FAILS on initial (all None) → PASSES on golden (FQ1-FQ4 in A6:A9)
    try:
        second_cycle_values = [ws.cell(row=r, column=1).value for r in range(6, 10)]
        expected_second = EXPECTED_CYCLE  # ['FQ1', 'FQ2', 'FQ3', 'FQ4']

        if second_cycle_values == expected_second:
            print(f"PASS: Component 2 — Second cycle A6:A9 = {second_cycle_values} (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 2 — Expected A6:A9 = {expected_second}, found {second_cycle_values}")
    except Exception as e:
        print(f"ERROR: Component 2 — Cannot read A6:A9: {e}")

    # Component 3: Third cycle A10:A13 = FQ1, FQ2, FQ3, FQ4 (0.35 points)
    # This FAILS on initial (all None) → PASSES on golden (FQ1-FQ4 in A10:A13)
    try:
        third_cycle_values = [ws.cell(row=r, column=1).value for r in range(10, 14)]
        expected_third = EXPECTED_CYCLE  # ['FQ1', 'FQ2', 'FQ3', 'FQ4']

        if third_cycle_values == expected_third:
            print(f"PASS: Component 3 — Third cycle A10:A13 = {third_cycle_values} (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 3 — Expected A10:A13 = {expected_third}, found {third_cycle_values}")
    except Exception as e:
        print(f"ERROR: Component 3 — Cannot read A10:A13: {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
