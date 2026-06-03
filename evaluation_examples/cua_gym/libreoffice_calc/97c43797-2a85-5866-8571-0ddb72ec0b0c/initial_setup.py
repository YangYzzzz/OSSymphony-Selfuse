"""
Initial Setup: Create FiscalData spreadsheet with empty A2:A13 for fill series task.
Task ID: calc_dop_fillseries_custom_052
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_dop_fillseries_custom_052'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: FiscalData ---
    ws = wb.active
    ws.title = 'FiscalData'

    # Headers in row 1
    headers = ['Fiscal Quarter', 'Revenue', 'Costs', 'Margin']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFFFF')
        cell.alignment = Alignment(horizontal='center')

    # Data for rows 2-13: A2:A13 MUST be empty (task is to fill them)
    # Revenue, Costs, Margin data for 12 fiscal quarters (3 years)
    financial_data = [
        # (Revenue, Costs, Margin%)
        (1_245_800, 832_400, 33.2),
        (1_378_500, 901_200, 34.6),
        (1_189_300, 798_600, 32.8),
        (1_456_700, 963_100, 33.9),
        (1_312_400, 875_800, 33.2),
        (1_498_200, 987_400, 34.1),
        (1_267_600, 841_300, 33.6),
        (1_534_900, 1_012_500, 34.0),
        (1_389_100, 921_700, 33.7),
        (1_556_300, 1_028_900, 33.9),
        (1_423_800, 943_200, 33.7),
        (1_612_400, 1_067_800, 33.8),
    ]

    for row_idx, (revenue, costs, margin) in enumerate(financial_data, 2):
        # A column (Fiscal Quarter): LEAVE EMPTY - this is what the task fills
        # B column: Revenue
        ws.cell(row=row_idx, column=2, value=revenue)
        ws.cell(row=row_idx, column=2).number_format = '#,##0'
        # C column: Costs
        ws.cell(row=row_idx, column=3, value=costs)
        ws.cell(row=row_idx, column=3).number_format = '#,##0'
        # D column: Margin
        ws.cell(row=row_idx, column=4, value=margin)
        ws.cell(row=row_idx, column=4).number_format = '0.0"%"'

    # Adjust column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheet: FiscalData')
    print('  Row 1: Headers (Fiscal Quarter, Revenue, Costs, Margin)')
    print('  A2:A13: EMPTY (to be filled by agent)')
    print('  B2:D13: Financial data for 12 quarters')


create_initial()
