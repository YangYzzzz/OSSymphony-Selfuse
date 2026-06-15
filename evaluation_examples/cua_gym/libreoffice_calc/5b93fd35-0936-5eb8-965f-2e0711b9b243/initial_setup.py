"""
Initial Setup: Sales order data with product catalog for profitability analysis
Task ID: calc_fin_profitability_vlookup_038
Domain: libreoffice_calc

Creates:
  - Sheet 'Orders': 99 rows of order data (A-D filled, E-I empty)
  - Sheet 'ProductCatalog': 39 product rows with cost data
"""

import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_profitability_vlookup_038'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ #
    # Sheet 1: Orders
    # ------------------------------------------------------------------ #
    ws_orders = wb.active
    ws_orders.title = 'Orders'

    # Headers (NOT bold — agent must make them bold)
    order_headers = ['OrderID', 'ProductCode', 'Qty', 'Unit Price',
                     'Unit Cost', 'Revenue', 'COGS', 'Gross Profit', 'Margin %']
    for col, h in enumerate(order_headers, 1):
        ws_orders.cell(row=1, column=col, value=h)

    # Product codes to use (must exist in ProductCatalog)
    product_codes = [
        'PRD-001', 'PRD-002', 'PRD-003', 'PRD-004', 'PRD-005',
        'PRD-006', 'PRD-007', 'PRD-008', 'PRD-009', 'PRD-010',
        'PRD-011', 'PRD-012', 'PRD-013', 'PRD-014', 'PRD-015',
        'PRD-016', 'PRD-017', 'PRD-018', 'PRD-019', 'PRD-020',
    ]

    # 99 order rows — only A (OrderID), B (ProductCode), C (Qty), D (UnitPrice)
    import random
    random.seed(42)

    unit_prices = {
        'PRD-001': 129.99, 'PRD-002': 249.50, 'PRD-003': 89.99,
        'PRD-004': 349.00, 'PRD-005': 59.99,  'PRD-006': 199.95,
        'PRD-007': 74.50,  'PRD-008': 449.00, 'PRD-009': 34.99,
        'PRD-010': 279.99, 'PRD-011': 149.95, 'PRD-012': 99.00,
        'PRD-013': 189.50, 'PRD-014': 329.00, 'PRD-015': 44.99,
        'PRD-016': 219.99, 'PRD-017': 109.75, 'PRD-018': 399.00,
        'PRD-019': 69.99,  'PRD-020': 289.50,
    }

    order_rows = []
    for i in range(1, 100):
        order_id = f'ORD-{i:04d}'
        pcode = random.choice(product_codes)
        qty = random.randint(1, 50)
        unit_price = unit_prices[pcode]
        # E (Unit Cost), F (Revenue), G (COGS), H (Gross Profit), I (Margin %) — all empty
        order_rows.append([order_id, pcode, qty, unit_price, None, None, None, None, None])

    for r, row_data in enumerate(order_rows, 2):
        for c, val in enumerate(row_data, 1):
            ws_orders.cell(row=r, column=c, value=val)

    # ------------------------------------------------------------------ #
    # Sheet 2: ProductCatalog
    # ------------------------------------------------------------------ #
    ws_cat = wb.create_sheet('ProductCatalog')

    catalog_headers = ['ProductCode', 'ProductName', 'UnitCost', 'Category']
    for col, h in enumerate(catalog_headers, 1):
        ws_cat.cell(row=1, column=col, value=h)

    catalog_data = [
        ('PRD-001', 'Wireless Bluetooth Headphones',   78.00,  'Electronics'),
        ('PRD-002', 'Mechanical Keyboard Pro',          148.00, 'Electronics'),
        ('PRD-003', 'Ergonomic Mouse Pad XL',           42.00,  'Accessories'),
        ('PRD-004', 'Ultra-Wide Curved Monitor',        210.00, 'Electronics'),
        ('PRD-005', 'USB-C Hub 7-Port',                 28.00,  'Accessories'),
        ('PRD-006', 'Noise-Cancelling Earbuds',         115.00, 'Electronics'),
        ('PRD-007', 'Laptop Stand Adjustable',          35.00,  'Accessories'),
        ('PRD-008', 'External SSD 2TB',                 260.00, 'Storage'),
        ('PRD-009', 'Cable Management Kit',             14.50,  'Accessories'),
        ('PRD-010', 'Webcam 4K 60fps',                  170.00, 'Electronics'),
        ('PRD-011', 'Smart LED Desk Lamp',              88.00,  'Lighting'),
        ('PRD-012', 'Portable Charger 20000mAh',        55.00,  'Electronics'),
        ('PRD-013', 'HDMI 2.1 Cable 3m',                85.00,  'Cables'),
        ('PRD-014', 'Graphics Tablet Drawing Pro',      195.00, 'Peripherals'),
        ('PRD-015', 'Screen Cleaning Kit',              18.00,  'Accessories'),
        ('PRD-016', 'Wireless Presenter Remote',        130.00, 'Accessories'),
        ('PRD-017', 'Microphone USB Condenser',         65.00,  'Audio'),
        ('PRD-018', 'NAS Storage Server 4-Bay',         245.00, 'Storage'),
        ('PRD-019', 'Webcam Privacy Cover 3-Pack',      28.00,  'Accessories'),
        ('PRD-020', 'Digital Whiteboard Stylus',        172.00, 'Peripherals'),
        ('PRD-021', 'Router Wi-Fi 6E Tri-Band',         180.00, 'Networking'),
        ('PRD-022', 'Switch 8-Port Gigabit',            55.00,  'Networking'),
        ('PRD-023', 'KVM Switch 4-Port',                95.00,  'Networking'),
        ('PRD-024', 'UPS 1500VA',                       210.00, 'Power'),
        ('PRD-025', 'Power Strip Surge Protector',      38.00,  'Power'),
        ('PRD-026', 'Monitor Arm Dual',                 105.00, 'Accessories'),
        ('PRD-027', 'Keyboard Wrist Rest Gel',          22.00,  'Accessories'),
        ('PRD-028', 'Fingerprint USB Security Key',     48.00,  'Security'),
        ('PRD-029', 'Smart Card Reader',                29.00,  'Security'),
        ('PRD-030', 'VoIP Phone Desktop',               145.00, 'Communications'),
        ('PRD-031', 'Conference Speaker 360°',          160.00, 'Audio'),
        ('PRD-032', 'Laser Pointer Professional',       25.00,  'Accessories'),
        ('PRD-033', 'Cable Tester Network',             40.00,  'Tools'),
        ('PRD-034', 'Patch Panel 24-Port',              90.00,  'Networking'),
        ('PRD-035', 'Rack Mount Server 1U Shelf',       75.00,  'Hardware'),
        ('PRD-036', 'Anti-Glare Screen Filter 27"',     42.00,  'Accessories'),
        ('PRD-037', 'Docking Station Thunderbolt 4',    185.00, 'Accessories'),
        ('PRD-038', 'Barcode Scanner 2D',               110.00, 'Peripherals'),
        ('PRD-039', 'Label Printer Thermal',            135.00, 'Peripherals'),
    ]

    for r, row_data in enumerate(catalog_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_cat.cell(row=r, column=c, value=val)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Orders sheet: 99 data rows (columns E-I empty, not computed)')
    print(f'  ProductCatalog sheet: 39 product rows')


create_initial()
