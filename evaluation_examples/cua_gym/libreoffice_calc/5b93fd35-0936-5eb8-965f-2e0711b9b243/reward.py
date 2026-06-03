"""
Reward Script: Enrich sales order data with VLOOKUP for profitability analysis
Task ID: calc_fin_profitability_vlookup_038
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: VLOOKUP formula in E2:E100 for Unit Cost        (0.25 pts)
  Component 2: Revenue/COGS/GrossProfit/Margin formulas F-I    (0.25 pts)
  Component 3: Currency number format on D-H, percentage on I  (0.20 pts)
  Component 4: Conditional formatting on I2:I100 < 0.20 red    (0.20 pts)
  Component 5: Row 1 headers are bold                          (0.10 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_profitability_vlookup_038'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify 'Orders' sheet exists
    if 'Orders' not in wb.sheetnames:
        print("FAIL: 'Orders' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Orders']

    # -----------------------------------------------------------------
    # Component 1: VLOOKUP formula in E2:E100 for Unit Cost (0.25 pts)
    # Each row E should contain: =IFERROR(VLOOKUP(B<n>,ProductCatalog.$A$2:$C$40,3,0),"N/A")
    # This FAILS on the initial file (E2:E100 are all None) and PASSES on golden.
    # -----------------------------------------------------------------
    try:
        vlookup_count = 0
        vlookup_expected = 99  # rows 2-100
        for row in range(2, 101):
            e_val = ws.cell(row=row, column=5).value  # Column E
            if e_val and isinstance(e_val, str):
                e_upper = e_val.upper().replace(' ', '')
                # Check for IFERROR+VLOOKUP with ProductCatalog reference
                if ('VLOOKUP' in e_upper and
                        'PRODUCTCATALOG' in e_upper and
                        'IFERROR' in e_upper and
                        '3,0' in e_upper.replace(' ', '')):
                    vlookup_count += 1

        if vlookup_count == vlookup_expected:
            print(f"PASS: Component 1 — VLOOKUP formula found in all E2:E100 ({vlookup_count}/{vlookup_expected}) (0.25 pts)")
            total_score += 0.25
        elif vlookup_count >= vlookup_expected // 2:
            print(f"PARTIAL: Component 1 — VLOOKUP formula found in {vlookup_count}/{vlookup_expected} rows (0.12 pts)")
            total_score += 0.12
        else:
            print(f"FAIL: Component 1 — VLOOKUP formula found in only {vlookup_count}/{vlookup_expected} rows (expected all 99)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------
    # Component 2: Revenue/COGS/GrossProfit/Margin formulas in F-I columns (0.25 pts)
    # F: =C*D (Revenue), G: =IF(E="N/A","N/A",C*E) (COGS),
    # H: =IF(G="N/A","N/A",F-G) (Gross Profit), I: =IF(H="N/A","N/A",H/F) (Margin %)
    # All these are None in initial file.
    # -----------------------------------------------------------------
    try:
        rev_count = 0    # Column F
        cogs_count = 0   # Column G
        gp_count = 0     # Column H
        margin_count = 0 # Column I

        for row in range(2, 101):
            f_val = ws.cell(row=row, column=6).value  # F: Revenue
            g_val = ws.cell(row=row, column=7).value  # G: COGS
            h_val = ws.cell(row=row, column=8).value  # H: Gross Profit
            i_val = ws.cell(row=row, column=9).value  # I: Margin %

            # Revenue: =C*D pattern
            if f_val and isinstance(f_val, str):
                f_upper = f_val.upper().replace(' ', '')
                if f_upper.startswith('=C') and '*D' in f_upper:
                    rev_count += 1

            # COGS: =IF(E...,"N/A",C*E) pattern
            if g_val and isinstance(g_val, str):
                g_upper = g_val.upper().replace(' ', '')
                if 'IF(' in g_upper and '"N/A"' in g_upper:
                    cogs_count += 1

            # Gross Profit: =IF(G...,"N/A",F-G) pattern
            if h_val and isinstance(h_val, str):
                h_upper = h_val.upper().replace(' ', '')
                if 'IF(' in h_upper and '"N/A"' in h_upper:
                    gp_count += 1

            # Margin %: =IF(H...,"N/A",H/F) pattern
            if i_val and isinstance(i_val, str):
                i_upper = i_val.upper().replace(' ', '')
                if 'IF(' in i_upper and '"N/A"' in i_upper:
                    margin_count += 1

        # All 4 formula types must be present in all 99 rows
        all_formulas_complete = (
            rev_count == 99 and
            cogs_count == 99 and
            gp_count == 99 and
            margin_count == 99
        )

        if all_formulas_complete:
            print(f"PASS: Component 2 — All formula columns F/G/H/I present in all 99 rows (0.25 pts)")
            total_score += 0.25
        else:
            # Partial credit: award points proportional to how many formula types are complete
            complete_cols = sum([rev_count == 99, cogs_count == 99, gp_count == 99, margin_count == 99])
            partial = round(complete_cols / 4 * 0.25, 3)
            if partial > 0:
                print(f"PARTIAL: Component 2 — {complete_cols}/4 formula columns complete "
                      f"(F:{rev_count}, G:{cogs_count}, H:{gp_count}, I:{margin_count}) ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 2 — Formula columns incomplete "
                      f"(F:{rev_count}, G:{cogs_count}, H:{gp_count}, I:{margin_count})/99 rows")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------
    # Component 3: Number formats (0.20 pts)
    # D2:H100 should be currency format ($#,##0.00)
    # I2:I100 should be percentage format (0.00%)
    # In the initial file, E-I are empty so no number format is applied.
    # -----------------------------------------------------------------
    try:
        currency_formats_ok = True
        percent_formats_ok = True
        currency_checked = 0
        percent_checked = 0

        # Sample every 10th row to check formats (rows 2, 12, 22, ..., 92)
        sample_rows = list(range(2, 101, 10))

        for row in sample_rows:
            # Check D column (Unit Price - should be currency)
            d_cell = ws.cell(row=row, column=4)
            d_fmt = d_cell.number_format or ''
            if '$' not in d_fmt and '#,##0' not in d_fmt:
                # D might already be formatted in initial, skip for currency check
                pass

            # Check E (Unit Cost), F (Revenue), G (COGS), H (Gross Profit) for currency
            for col in range(5, 9):  # E, F, G, H
                cell = ws.cell(row=row, column=col)
                fmt = cell.number_format or ''
                if '$' not in fmt and '#,##0' not in fmt:
                    currency_formats_ok = False
                currency_checked += 1

            # Check I (Margin %) for percentage
            i_cell = ws.cell(row=row, column=9)
            i_fmt = i_cell.number_format or ''
            if '%' not in i_fmt:
                percent_formats_ok = False
            percent_checked += 1

        if currency_formats_ok and percent_checked > 0 and percent_formats_ok:
            print(f"PASS: Component 3 — Currency format on E-H columns and percentage on I column (0.20 pts)")
            total_score += 0.20
        elif currency_formats_ok:
            print(f"PARTIAL: Component 3 — Currency format OK but percentage format on I missing/incorrect (0.10 pts)")
            total_score += 0.10
        elif percent_formats_ok:
            print(f"PARTIAL: Component 3 — Percentage format on I OK but currency format on E-H missing (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — Number formats not applied correctly (currency_ok={currency_formats_ok}, percent_ok={percent_formats_ok})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------
    # Component 4: Conditional formatting on I2:I100 (0.20 pts)
    # Rule: cellIs < 0.20 with red fill (FFFF0000)
    # Initial file has no conditional formatting.
    # -----------------------------------------------------------------
    try:
        cf_found = False
        cf_correct_range = False
        cf_correct_rule = False
        cf_correct_color = False

        for cf_range in ws.conditional_formatting:
            cf_range_str = str(cf_range)
            cf_list = ws.conditional_formatting[cf_range]

            # Check if CF applies to I2:I100 (or a superset like I:I)
            if 'I2' in cf_range_str and 'I100' in cf_range_str:
                cf_correct_range = True
            elif 'I' in cf_range_str:
                cf_correct_range = True  # Accept column-level CF too

            for rule in cf_list:
                cf_found = True
                rule_type = rule.type
                rule_op = rule.operator
                rule_formula = rule.formula

                # Check cellIs < 0.2 rule
                if (rule_type == 'cellIs' and
                        rule_op == 'lessThan' and
                        rule_formula and
                        any('0.2' in str(f) for f in rule_formula)):
                    cf_correct_rule = True

                # Check for red fill in the DXF style
                try:
                    dxf = rule.dxf
                    if dxf and dxf.fill:
                        fill = dxf.fill
                        # Check fgColor for red (FFFF0000 or FF0000 or similar)
                        try:
                            fg_rgb = fill.fgColor.rgb if fill.fgColor else ''
                            if fg_rgb and 'FF0000' in fg_rgb.upper():
                                cf_correct_color = True
                        except Exception:
                            pass
                        # Also check bgColor
                        try:
                            bg_rgb = fill.bgColor.rgb if fill.bgColor else ''
                            if bg_rgb and 'FF0000' in bg_rgb.upper():
                                cf_correct_color = True
                        except Exception:
                            pass
                except Exception:
                    pass

        if cf_correct_range and cf_correct_rule and cf_correct_color:
            print(f"PASS: Component 4 — Conditional formatting on I2:I100, cellIs<0.2 with red fill (0.20 pts)")
            total_score += 0.20
        elif cf_found and cf_correct_rule:
            print(f"PARTIAL: Component 4 — CF rule correct (cellIs<0.2) but range or color may differ (0.10 pts)")
            total_score += 0.10
        elif cf_found:
            print(f"PARTIAL: Component 4 — CF exists but rule/range/color incorrect "
                  f"(range={cf_correct_range}, rule={cf_correct_rule}, color={cf_correct_color}) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4 — No conditional formatting found on Orders sheet")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -----------------------------------------------------------------
    # Component 5: Row 1 headers are bold (0.10 pts)
    # Initial file has non-bold headers.
    # -----------------------------------------------------------------
    try:
        bold_count = 0
        for col in range(1, 10):  # A through I
            cell = ws.cell(row=1, column=col)
            if cell.font and cell.font.bold:
                bold_count += 1

        if bold_count == 9:
            print(f"PASS: Component 5 — All 9 header cells in row 1 are bold (0.10 pts)")
            total_score += 0.10
        elif bold_count >= 5:
            print(f"PARTIAL: Component 5 — {bold_count}/9 header cells are bold (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — Only {bold_count}/9 header cells are bold (expected all 9)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
