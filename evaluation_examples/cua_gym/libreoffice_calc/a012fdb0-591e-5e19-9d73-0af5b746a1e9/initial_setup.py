"""
Initial Setup: Weekly template with B2:B52 filled with numeric data and formatting
Task ID: calc_cop_clear_005
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_cop_clear_005'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'WeeklyTemplate'

    # Colors for alternating rows
    WHITE = 'FFFFFFFF'
    LIGHT_GRAY = 'FFD9D9D9'

    # Border style
    thin = Side(style='thin', color='000000')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Column A header
    ws['A1'] = 'Day'
    ws['A1'].font = Font(bold=True, name='Calibri', size=11)
    ws['A1'].fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    ws['A1'].border = thin_border

    # Column B header
    ws['B1'] = 'Units Sold'
    ws['B1'].font = Font(bold=True, name='Calibri', size=11)
    ws['B1'].fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    ws['B1'].border = thin_border

    # Column C header
    ws['C1'] = 'Revenue ($)'
    ws['C1'].font = Font(bold=True, name='Calibri', size=11)
    ws['C1'].fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    ws['C1'].border = thin_border

    # Column D header
    ws['D1'] = 'Returns'
    ws['D1'].font = Font(bold=True, name='Calibri', size=11)
    ws['D1'].fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    ws['D1'].border = thin_border

    # Realistic day labels and data for 51 rows
    # 7 weeks + a couple extra days
    day_names = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    units_data = [
        142, 178, 203, 165, 244, 312, 289,   # week 1
        156, 191, 217, 183, 261, 298, 275,   # week 2
        134, 168, 195, 172, 238, 321, 301,   # week 3
        149, 185, 208, 177, 255, 307, 284,   # week 4
        163, 197, 221, 188, 267, 315, 293,   # week 5
        141, 174, 199, 169, 247, 302, 278,   # week 6
        158, 193, 215, 181, 259, 311, 288,   # week 7
        147, 179, 204,                        # 3 extra days
    ]
    revenue_data = [
        2840.50, 3560.00, 4060.00, 3300.00, 4880.00, 6240.00, 5780.00,
        3120.00, 3820.00, 4340.00, 3660.00, 5220.00, 5960.00, 5500.00,
        2680.00, 3360.00, 3900.00, 3440.00, 4760.00, 6420.00, 6020.00,
        2980.00, 3700.00, 4160.00, 3540.00, 5100.00, 6140.00, 5680.00,
        3260.00, 3940.00, 4420.00, 3760.00, 5340.00, 6300.00, 5860.00,
        2820.00, 3480.00, 3980.00, 3380.00, 4940.00, 6040.00, 5560.00,
        3160.00, 3860.00, 4300.00, 3620.00, 5180.00, 6220.00, 5760.00,
        2940.00, 3580.00, 4080.00,
    ]
    returns_data = [
        5, 7, 9, 6, 11, 14, 12,
        6, 8, 10, 7, 12, 13, 11,
        4, 7, 8, 7, 10, 15, 13,
        5, 8, 9, 7, 11, 14, 12,
        6, 9, 10, 8, 12, 14, 13,
        5, 7, 9, 6, 11, 13, 12,
        6, 8, 10, 7, 12, 14, 12,
        5, 7, 9,
    ]

    for i in range(51):
        row = i + 2  # rows 2 to 52
        week_num = i // 7 + 1
        day_idx = i % 7
        label = f'W{week_num}-{day_names[day_idx]}'

        # Alternating background colors (white for even, light-gray for odd)
        bg_color = WHITE if (i % 2 == 0) else LIGHT_GRAY

        # Every 5th row (0-indexed: row 4, 9, 14, ...) is bold
        is_bold = ((i + 1) % 5 == 0)

        fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        font_normal = Font(name='Calibri', size=10, bold=is_bold)
        font_b = Font(name='Calibri', size=10, bold=is_bold)

        # Column A: labels (no task modification)
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = font_normal
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='left')

        # Column B: numeric data (will be cleared in golden)
        ws.cell(row=row, column=2, value=units_data[i])
        ws.cell(row=row, column=2).font = font_b
        ws.cell(row=row, column=2).fill = fill
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')

        # Column C: revenue (unchanged by task)
        ws.cell(row=row, column=3, value=revenue_data[i])
        ws.cell(row=row, column=3).font = Font(name='Calibri', size=10)
        ws.cell(row=row, column=3).fill = fill
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=3).number_format = '#,##0.00'

        # Column D: returns (unchanged by task)
        ws.cell(row=row, column=4, value=returns_data[i])
        ws.cell(row=row, column=4).font = Font(name='Calibri', size=10)
        ws.cell(row=row, column=4).fill = fill
        ws.cell(row=row, column=4).border = thin_border
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')

    # Set column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

create_initial()
