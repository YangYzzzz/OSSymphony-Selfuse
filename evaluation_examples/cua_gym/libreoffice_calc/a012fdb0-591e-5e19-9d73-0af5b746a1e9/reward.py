"""
Reward Script: Clear contents only from column B (B2:B52) while leaving all formatting intact
Task ID: calc_cop_clear_005
Domain: libreoffice_calc
Scoring:
  - Component 1: All B2:B52 values are cleared to None (0.5 points)
  - Component 2: Cleared cells preserve alternating fill colors (compound: value=None AND fill=correct) (0.3 points)
  - Component 3: Cleared cells preserve bold formatting on designated rows (compound: value=None AND bold=correct) (0.2 points)

NOTE: Components 2 and 3 are compound checks — they require BOTH the value-clearing AND
formatting preservation to be correct. This ensures the initial file (which has values)
scores 0.0, since it fails the "value cleared" part of the compound.
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_cop_clear_005'
SHEET_NAME = 'WeeklyTemplate'

# Rows that should have bold formatting (every 5th row starting from row 6)
BOLD_ROWS = {6, 11, 16, 21, 26, 31, 36, 41, 46, 51}

# Expected fill pattern for B2:B52:
# Index 0 (B2): FFFFFFFF (white), Index 1 (B3): FFD9D9D9 (gray), alternating
EXPECTED_FILLS = {}
for r in range(2, 53):
    idx = r - 2
    if idx % 2 == 0:
        EXPECTED_FILLS[r] = 'FFFFFFFF'
    else:
        EXPECTED_FILLS[r] = 'FFD9D9D9'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Component 1: All B2:B52 values are cleared (0.5 points)
    # This FAILS on initial (values 142-179 are present) and PASSES on golden (all None).
    try:
        cleared_count = 0
        not_cleared = []
        for row in range(2, 53):
            val = ws.cell(row=row, column=2).value
            if val is None:
                cleared_count += 1
            else:
                not_cleared.append(f"B{row}={repr(val)}")

        if cleared_count == 51:
            print(f"PASS: Component 1 — All 51 cells B2:B52 are cleared (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — Only {cleared_count}/51 cells cleared. Not cleared: {not_cleared[:5]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Cleared cells preserve alternating fill colors (0.3 points)
    # Compound check: cell must be CLEARED (value=None) AND have correct fill.
    # This FAILS on initial because values are not None (even though fills are correct).
    # This PASSES on golden because values are None AND fills are preserved.
    try:
        compound_fill_ok = 0
        compound_fill_broken = []
        for row in range(2, 53):
            cell = ws.cell(row=row, column=2)
            # Both conditions must hold: value cleared AND fill preserved
            val_cleared = (cell.value is None)
            try:
                actual_fill = cell.fill.fgColor.rgb
            except Exception:
                actual_fill = None
            expected = EXPECTED_FILLS[row]
            fill_ok = (actual_fill == expected)

            if val_cleared and fill_ok:
                compound_fill_ok += 1
            else:
                if not val_cleared:
                    compound_fill_broken.append(f"B{row}: value not cleared (={repr(cell.value)})")
                elif not fill_ok:
                    compound_fill_broken.append(f"B{row}: fill={actual_fill}, expected={expected}")

        if compound_fill_ok == 51:
            print(f"PASS: Component 2 — All 51 cells cleared AND fill preserved (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — Only {compound_fill_ok}/51 cells cleared+fill correct. Issues: {compound_fill_broken[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Cleared cells preserve bold formatting on designated rows (0.2 points)
    # Compound check: the 10 designated bold rows must be CLEARED AND have bold=True.
    # Non-bold rows must be CLEARED AND have bold=False.
    # This FAILS on initial because values are not cleared.
    # This PASSES on golden because values are cleared AND bold is preserved.
    try:
        bold_compound_ok = 0
        bold_compound_broken = []

        for row in range(2, 53):
            cell = ws.cell(row=row, column=2)
            val_cleared = (cell.value is None)
            expected_bold = (row in BOLD_ROWS)
            actual_bold = cell.font.bold if cell.font else False

            if val_cleared and (actual_bold == expected_bold):
                bold_compound_ok += 1
            else:
                if not val_cleared:
                    bold_compound_broken.append(f"B{row}: value not cleared")
                else:
                    bold_compound_broken.append(f"B{row}: expected bold={expected_bold}, got bold={actual_bold}")

        if bold_compound_ok == 51:
            print(f"PASS: Component 3 — All 51 cells cleared AND bold formatting preserved (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — Only {bold_compound_ok}/51 cells cleared+bold correct. Issues: {bold_compound_broken[:3]}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
