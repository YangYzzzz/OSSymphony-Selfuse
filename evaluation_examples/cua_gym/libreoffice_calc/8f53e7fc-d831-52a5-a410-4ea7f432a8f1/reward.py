"""
FINAL REWARD SCRIPT - SUCCESS
Task: I want a table with column headers ("Region" and "Customers") in a new sheet named "Regional_Summary" showing total customers per region.
Generated: 2025-11-24 07:43:57
Status: success
Model: o3
Total Steps: 1
"""

import openpyxl
import os
import math


def verify_regional_summary(file_path: str) -> float:
    """Verify that the workbook contains a sheet named 'Regional_Summary' with
    a two-column table (Region, Customers) showing correct customer counts per
    region based on data in the original 'Data' sheet.

    Scoring (progressive up to 1.0):
        0.25 – Sheet 'Regional_Summary' exists
        0.25 – Headers exactly 'Region' and 'Customers'
        0.25 – All regions that appear in 'Data' are present in the summary
        0.25 – Customer counts per region are correct
    """

    print(f"Verifying file: {file_path}")
    total_score = 0.0
    max_score = 1.0

    # --------------------------------------------------
    # 1. Load workbook
    # --------------------------------------------------
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        print("✓ Workbook loaded")
    except Exception as e:
        print(f"✗ Failed to load workbook: {e}")
        return 0.0  # Cannot continue without the workbook

    # --------------------------------------------------
    # 2. Gather expected counts from 'Data' sheet
    # --------------------------------------------------
    if 'Data' not in wb.sheetnames:
        print("✗ 'Data' sheet missing; cannot verify task")
        return 0.0

    data_sheet = wb['Data']
    expected_counts = {}

    for idx, row in enumerate(data_sheet.iter_rows(values_only=True), start=1):
        if idx == 1:  # skip header
            continue
        region = row[0]
        if region is None:
            continue
        expected_counts[region] = expected_counts.get(region, 0) + 1

    print(f"Expected counts per region from 'Data': {expected_counts}")

    # --------------------------------------------------
    # 3. Verify 'Regional_Summary' sheet existence
    # --------------------------------------------------
    if 'Regional_Summary' in wb.sheetnames:
        print("✓ 'Regional_Summary' sheet found (0.25 points)")
        total_score += 0.25
        summary_sheet = wb['Regional_Summary']
    else:
        print("✗ 'Regional_Summary' sheet not found")
        return total_score  # No further checks possible

    # --------------------------------------------------
    # 4. Verify headers and read summary counts
    # --------------------------------------------------
    rows = list(summary_sheet.iter_rows(values_only=True))
    headers_ok = False
    summary_counts = {}

    if rows:
        header_row = rows[0]
        if header_row and len(header_row) >= 2:
            h1 = str(header_row[0]).strip().lower() if header_row[0] is not None else ""
            h2 = str(header_row[1]).strip().lower() if header_row[1] is not None else ""
            if h1 == 'region' and h2 == 'customers':
                headers_ok = True
                print("✓ Headers 'Region' and 'Customers' found (0.25 points)")
                total_score += 0.25
            else:
                print(f"✗ Headers mismatch. Found: {header_row}")
        else:
            print("✗ Header row incomplete or empty")
    else:
        print("✗ 'Regional_Summary' sheet is empty")

    # Read data rows (if any) into summary_counts
    if len(rows) > 1:
        for r in rows[1:]:
            region = r[0]
            customers_val = r[1]
            if region is None:
                continue
            summary_counts[region] = customers_val
        print(f"Summary counts read: {summary_counts}")
    else:
        print("✗ No data rows found in summary sheet")

    # --------------------------------------------------
    # 5. Check that all regions are present
    # --------------------------------------------------
    if expected_counts and summary_counts:
        missing = [reg for reg in expected_counts if reg not in summary_counts]
        extra = [reg for reg in summary_counts if reg not in expected_counts]

        if not missing:
            print("✓ All regions from data present in summary (0.25 points)")
            total_score += 0.25
        else:
            print(f"✗ Missing regions in summary: {missing}")

        if extra:
            print(f"Note: Extra regions present in summary: {extra}")
    else:
        print("✗ Unable to verify region list completeness")

    # --------------------------------------------------
    # 6. Verify customer counts correctness
    # --------------------------------------------------
    counts_correct = True
    incorrect_details = []

    for reg, expected in expected_counts.items():
        if reg in summary_counts:
            actual = summary_counts[reg]

            # Convert to integer if possible
            if isinstance(actual, str) and actual.strip().isdigit():
                actual = int(actual.strip())
            if isinstance(actual, float) and math.isclose(actual, round(actual)):
                actual = int(round(actual))

            if actual != expected:
                counts_correct = False
                incorrect_details.append((reg, expected, actual))
        else:
            counts_correct = False
            incorrect_details.append((reg, expected, None))

    if summary_counts:
        if counts_correct:
            print("✓ Customer counts correct for all regions (0.25 points)")
            total_score += 0.25
        else:
            print(f"✗ Incorrect counts found: {incorrect_details}")

    # --------------------------------------------------
    # 7. Final score
    # --------------------------------------------------
    final_score = round(min(total_score, max_score), 2)
    print(f"Total score: {final_score}/{max_score}")
    return final_score


if __name__ == "__main__":
    workbook_path = "/home/user/i_want_a_table_with_column_headers_region_and_customers_in_a_new_sheet_named_regional_summary_showin.xlsx"
    reward = verify_regional_summary(workbook_path)
    print(f"REWARD: {reward}")
