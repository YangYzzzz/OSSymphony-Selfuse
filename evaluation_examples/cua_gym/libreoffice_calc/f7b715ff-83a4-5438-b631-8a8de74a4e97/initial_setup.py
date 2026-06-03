"""
Initial Setup: Create a Performance Ranking spreadsheet with employee data
Task ID: calc_gcv_055
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_055'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Performance_Ranking"

    # Headers
    headers = ['Employee Name', 'Department', 'Score']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    white_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 29 employees with realistic names, departments, and scores ranging 35-98
    employees = [
        ('Sarah Chen', 'Engineering', 92),
        ('Marcus Johnson', 'Marketing', 78),
        ('Elena Rodriguez', 'Finance', 85),
        ('David Kim', 'Engineering', 45),
        ('Priya Sharma', 'HR', 67),
        ('James O\'Brien', 'Sales', 98),
        ('Aisha Patel', 'Engineering', 73),
        ('Robert Thompson', 'Finance', 56),
        ('Lin Wei', 'Marketing', 88),
        ('Catherine Dubois', 'Sales', 41),
        ('Michael Brown', 'HR', 62),
        ('Yuki Tanaka', 'Engineering', 94),
        ('Sofia Martinez', 'Marketing', 35),
        ('William Harris', 'Finance', 81),
        ('Fatima Al-Rashid', 'Sales', 76),
        ('Daniel Lee', 'Engineering', 58),
        ('Olivia Turner', 'HR', 90),
        ('Ahmed Hassan', 'Marketing', 47),
        ('Rachel Green', 'Finance', 83),
        ('Thomas Anderson', 'Sales', 69),
        ('Maya Singh', 'Engineering', 96),
        ('Patrick Murphy', 'HR', 52),
        ('Isabella Costa', 'Marketing', 71),
        ('Hiroshi Yamamoto', 'Finance', 87),
        ('Emma Wilson', 'Sales', 39),
        ('Lucas Garcia', 'Engineering', 65),
        ('Natasha Volkov', 'HR', 79),
        ('Benjamin Scott', 'Finance', 43),
        ('Zara Ahmed', 'Marketing', 91),
    ]

    for r, (name, dept, score) in enumerate(employees, 2):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=dept)
        ws.cell(row=r, column=3, value=score)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10

    # No conditional formatting - that is the task for the agent

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
