"""
Reward Script: Year-over-year % change rows with conditional formatting
Task ID: osworld_calc_annual_pct_change_006
Domain: libreoffice_calc

Task: For a 5-year financial table (2019-2023) with 3 asset categories (CA, FA, OA),
insert 4 % change rows (one per year transition) with correct percentage formulas,
and apply conditional formatting: green for positive, red for negative, yellow for zero.

Scoring Rubric:
  Component 1: 4 % change rows present with correct labels          — 0.30 pts
  Component 2: % change formulas correct in B-D for all 4 rows     — 0.40 pts
  Component 3: Conditional formatting (green/red/yellow) applied   — 0.30 pts
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_annual_pct_change_006'

# Expected % change row labels
EXPECTED_LABELS = [
    '% Change 2019\u21922020',
    '% Change 2020\u21922021',
    '% Change 2021\u21922022',
    '% Change 2022\u21922023',
]

# Expected year rows in order (year values)
EXPECTED_YEARS = [2019, 2020, 2021, 2022, 2023]

# Expected % change row indices in the golden file (1-based)
PCT_ROWS = [3, 5, 7, 9]

# Expected conditional formatting colors (ARGB)
COLOR_GREEN  = 'FF00FF00'
COLOR_RED    = 'FFFF0000'
COLOR_YELLOW = 'FFFFFF00'


def normalize_label(label):
    """Normalize a label string for comparison (strip, lower, normalize arrow chars)."""
    if label is None:
        return ''
    s = str(label).strip()
    # Normalize various arrow representations
    s = s.replace('\u2192', '->').replace('\u21a6', '->')
    s = s.lower()
    return s


def labels_match(actual, expected):
    """Check if two % change labels refer to the same year transition."""
    a = normalize_label(actual)
    e = normalize_label(expected)
    if a == e:
        return True
    # Also accept partial match: both years must appear in order
    for arrow in ['\u2192', '->', '\u21a6', ' to ']:
        if arrow in e.replace('->', arrow):
            parts = e.replace('\u2192', '->').split('->')
            if len(parts) == 2:
                yr1 = parts[0].replace('% change ', '').strip()
                yr2 = parts[1].strip()
                return yr1 in a and yr2 in a
    return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # -----------------------------------------------------------------------
    # Component 1: 4 % change rows inserted with correct labels (0.30 points)
    # -----------------------------------------------------------------------
    # Expect 10 rows: interleaved year rows and % change rows
    # Row pattern: [header, 2019, %19->20, 2020, %20->21, 2021, %21->22, 2022, %22->23, 2023]
    try:
        max_row = ws.max_row
        print(f"INFO: Spreadsheet has {max_row} rows (expected 10 for fully completed task)")

        found_pct_rows = 0
        found_labels = []

        if max_row == 10:
            # Check that rows 3, 5, 7, 9 contain % change labels
            for pct_row_idx, expected_label in zip(PCT_ROWS, EXPECTED_LABELS):
                label_cell = ws.cell(row=pct_row_idx, column=1).value
                found_labels.append(label_cell)
                if label_cell is not None and '% Change' in str(label_cell) or \
                   label_cell is not None and '%' in str(label_cell):
                    found_pct_rows += 1

            if found_pct_rows == 4:
                print(f"PASS: Component 1 — 4 % change rows found at rows 3,5,7,9 with labels: {found_labels} (0.30 pts)")
                total_score += 0.30
            else:
                print(f"FAIL: Component 1 — only {found_pct_rows}/4 % change rows detected, labels={found_labels}")
        else:
            # Row count wrong — check if any % change rows exist at all
            pct_count = 0
            for row_idx in range(2, max_row + 1):
                val = ws.cell(row=row_idx, column=1).value
                if val is not None and '%' in str(val):
                    pct_count += 1
            print(f"FAIL: Component 1 — expected 10 rows, found {max_row}. % change labels detected: {pct_count}/4")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: % change formulas correct in B, C, D for all 4 rows (0.40 pts)
    # Each row earns 0.10 points; formula must reference correct adjacent cells
    # Expected formulas pattern: =(B{next_year_row} - B{prev_year_row}) / B{prev_year_row}
    # In golden: rows 3,5,7,9 are % change; rows 2,4,6,8,10 are data years
    # -----------------------------------------------------------------------
    try:
        # Map of % change row -> (prev_year_row, next_year_row)
        pct_row_data_map = {
            3: (2, 4),
            5: (4, 6),
            7: (6, 8),
            9: (8, 10),
        }

        formula_score = 0.0
        formulas_ok_per_row = {}

        if ws.max_row == 10:
            for pct_row, (prev_row, next_row) in pct_row_data_map.items():
                cols_ok = 0
                for col_idx, col_letter in [(2, 'B'), (3, 'C'), (4, 'D')]:
                    cell_val = ws.cell(row=pct_row, column=col_idx).value
                    if cell_val is not None and isinstance(cell_val, str) and cell_val.startswith('='):
                        # Normalize formula: remove spaces, uppercase
                        formula_norm = cell_val.replace(' ', '').upper()
                        # Check pattern: (Xcol{next} - Xcol{prev}) / Xcol{prev}
                        # Expected: =(B{next}-B{prev})/B{prev}
                        expected_pattern = f'=({col_letter}{next_row}-{col_letter}{prev_row})/{col_letter}{prev_row}'
                        if formula_norm == expected_pattern.upper():
                            cols_ok += 1
                        else:
                            print(f"  FAIL formula at {col_letter}{pct_row}: got {repr(cell_val)}, expected pattern {expected_pattern}")
                    else:
                        print(f"  FAIL formula at col {col_letter} row {pct_row}: not a formula, got {repr(cell_val)}")

                formulas_ok_per_row[pct_row] = cols_ok
                if cols_ok == 3:
                    formula_score += 0.10
                    print(f"  PASS: % change row {pct_row} — all 3 formulas correct (0.10 pts)")
                else:
                    print(f"  PARTIAL: % change row {pct_row} — {cols_ok}/3 formulas correct")

            if formula_score == 0.40:
                print(f"PASS: Component 2 — all 4 % change rows have correct formulas ({formula_score:.2f} pts)")
            elif formula_score > 0.0:
                print(f"PARTIAL: Component 2 — {formula_score:.2f}/0.40 pts — some formula rows correct")
            else:
                print(f"FAIL: Component 2 — no correct formulas found")
            total_score += formula_score
        else:
            print(f"FAIL: Component 2 — cannot check formulas, row count is {ws.max_row} (expected 10)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Conditional formatting (green/>0, red/<0, yellow/=0) (0.30 pts)
    # Each of the 4 % change row ranges must have all 3 rules applied
    # Expected ranges: B3:D3, B5:D5, B7:D7, B9:D9
    # -----------------------------------------------------------------------
    try:
        expected_cf_ranges = ['B3:D3', 'B5:D5', 'B7:D7', 'B9:D9']
        cf_rules_map = ws.conditional_formatting._cf_rules

        # Build set of range strings actually present
        cf_range_strs = set()
        for cf_obj in cf_rules_map.keys():
            cf_range_strs.add(str(cf_obj))

        ranges_with_correct_cf = 0
        for expected_range in expected_cf_ranges:
            # Find matching range: cf_obj key may be a ConditionalFormatting object
            # whose str() representation includes the range. Match by sqref or repr.
            matching_rules = None
            for cf_obj, rules in cf_rules_map.items():
                # Try direct string match, or check if expected_range is in the repr
                cf_str = str(cf_obj)
                if cf_str == expected_range or expected_range in cf_str:
                    matching_rules = rules
                    break
                # Also try sqref attribute if available
                sqref = getattr(cf_obj, 'sqref', None)
                if sqref is not None and str(sqref) == expected_range:
                    matching_rules = rules
                    break

            if matching_rules is None:
                print(f"  FAIL CF: range {expected_range} not found. Available ranges: {list(cf_range_strs)}")
                continue

            # Check for 3 rules: greaterThan/green, lessThan/red, equal/yellow
            has_green = False
            has_red = False
            has_yellow = False

            for rule in matching_rules:
                if rule.type != 'cellIs':
                    continue
                dxf = getattr(rule, 'dxf', None)
                if dxf is None:
                    continue
                fill = getattr(dxf, 'fill', None)
                if fill is None:
                    continue
                try:
                    color_rgb = fill.fgColor.rgb
                except Exception:
                    continue

                op = getattr(rule, 'operator', '')
                formula = getattr(rule, 'formula', [])

                if op == 'greaterThan' and formula == ['0'] and color_rgb == COLOR_GREEN:
                    has_green = True
                elif op == 'lessThan' and formula == ['0'] and color_rgb == COLOR_RED:
                    has_red = True
                elif op == 'equal' and formula == ['0'] and color_rgb == COLOR_YELLOW:
                    has_yellow = True

            if has_green and has_red and has_yellow:
                ranges_with_correct_cf += 1
                print(f"  PASS CF: range {expected_range} has all 3 conditional formatting rules (green/red/yellow)")
            else:
                print(f"  FAIL CF: range {expected_range} — green={has_green}, red={has_red}, yellow={has_yellow}")

        cf_score = round(ranges_with_correct_cf * 0.075, 3)  # 4 ranges * 0.075 = 0.30
        if ranges_with_correct_cf == 4:
            print(f"PASS: Component 3 — all 4 % change ranges have correct CF rules (0.30 pts)")
            total_score += 0.30
        elif ranges_with_correct_cf > 0:
            print(f"PARTIAL: Component 3 — {ranges_with_correct_cf}/4 ranges have correct CF, adding {cf_score:.3f} pts")
            total_score += cf_score
        else:
            print(f"FAIL: Component 3 — no % change ranges have correct conditional formatting")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the VM environment
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
