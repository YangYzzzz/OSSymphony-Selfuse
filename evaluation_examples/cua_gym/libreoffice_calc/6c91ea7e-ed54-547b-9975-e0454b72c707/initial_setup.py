"""
Initial Setup: Year-over-year asset analysis spreadsheet
Task ID: osworld_calc_annual_pct_change_006
Domain: libreoffice_calc

Creates a spreadsheet with 5 years of data (2019-2023) for three asset categories
(Current Assets, Fixed Assets, Other Assets) WITHOUT any % change rows or
conditional formatting. The agent must add these.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_annual_pct_change_006'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Annual Assets ---
    ws = wb.active
    ws.title = 'Annual Assets'

    # Headers
    headers = ['Year', 'CA (Current Assets)', 'FA (Fixed Assets)', 'OA (Other Assets)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # 5 years of realistic financial data (2019-2023)
    # Values in thousands USD — realistic asset values for a mid-size company
    data = [
        [2019, 482500, 1254000, 87300],
        [2020, 431200, 1198500, 79800],   # COVID impact: CA/FA down
        [2021, 518700, 1312000, 93400],   # Recovery
        [2022, 574300, 1389500, 105600],  # Growth
        [2023, 621800, 1456200, 112400],  # Continued growth
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22

    # Row 1 height for header
    ws.row_dimensions[1].height = 20

    # NOTE: No % change rows, no conditional formatting — agent must add these

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
