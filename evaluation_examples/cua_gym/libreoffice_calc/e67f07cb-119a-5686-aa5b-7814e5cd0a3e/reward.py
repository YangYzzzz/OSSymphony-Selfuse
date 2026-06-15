"""
Reward Script: Read sales receipts and enter them into daily_sales.xlsx with EOD report
Task ID: osworld_multi_apps_receipt_to_calc_015
Domain: libreoffice_calc
Scoring:
  Component 1: Transactions sheet populated with all 12 entries, valid payment methods and amounts (0.35)
  Component 2: EOD Report SUMIF formulas by payment method + Grand Total (0.25)
  Component 3: EOD Report summary statistics formulas (AVERAGE, COUNTA) (0.15)
  Component 4: EOD Report COUNTIFS for hourly transaction frequency (0.15)
  Component 5: Column chart of hourly transaction volume present in EOD Report (0.10)
Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_receipt_to_calc_015'

# Known receipt data for ground-truth verification
EXPECTED_PAYMENT_METHODS = {'Cash', 'Card', 'Mobile'}


def scan_eod_formulas(ws_eod):
    """Scan all rows of EOD Report sheet for formula types.
    Returns a dict of formula category -> count found.
    """
    sumif_cash_count = 0
    sumif_card_count = 0
    sumif_mobile_count = 0
    grand_total_count = 0
    count_formula_count = 0
    average_formula_count = 0
    countifs_formula_count = 0

    for row_num in range(1, 35):
        cell_val = ws_eod.cell(row=row_num, column=2).value
        if not isinstance(cell_val, str):
            continue
        val_upper = cell_val.upper()

        if 'SUMIF' in val_upper and 'CASH' in val_upper:
            sumif_cash_count += 1
        elif 'SUMIF' in val_upper and 'CARD' in val_upper:
            sumif_card_count += 1
        elif 'SUMIF' in val_upper and 'MOBILE' in val_upper:
            sumif_mobile_count += 1
        elif ('=SUM(' in val_upper and 'SUMIF' not in val_upper
              and re.search(r'=SUM\(B\d+:B\d+\)', cell_val, re.IGNORECASE)):
            grand_total_count += 1
        elif (('COUNTA' in val_upper or 'COUNT(' in val_upper)
              and 'TRANSACTIONS' in val_upper):
            count_formula_count += 1
        elif 'AVERAGE' in val_upper and 'TRANSACTIONS' in val_upper:
            average_formula_count += 1

        if 'COUNTIFS' in val_upper and 'TRANSACTIONS' in val_upper:
            countifs_formula_count += 1

    return {
        'sumif_cash': sumif_cash_count,
        'sumif_card': sumif_card_count,
        'sumif_mobile': sumif_mobile_count,
        'grand_total': grand_total_count,
        'count_stat': count_formula_count,
        'average_stat': average_formula_count,
        'countifs': countifs_formula_count,
    }


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook — precondition gate
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Required sheets gate
    if 'Transactions' not in wb.sheetnames or 'EOD Report' not in wb.sheetnames:
        print(f"CRITICAL: Required sheets not found. Found: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws_tx = wb['Transactions']
    ws_eod = wb['EOD Report']

    # -------------------------------------------------------------------------
    # Component 1a: Transactions sheet row count (0.20 points)
    # The Transactions sheet should have 12 data rows (R-001 through R-012)
    # in rows 2-13 — distinct from initial state (which has only placeholder row)
    # -------------------------------------------------------------------------
    try:
        valid_rows = 0
        payment_methods_found = set()
        valid_amounts = []

        for row_num in range(2, 14):  # rows 2-13
            receipt_id = ws_tx.cell(row=row_num, column=1).value
            payment = ws_tx.cell(row=row_num, column=4).value
            amount = ws_tx.cell(row=row_num, column=6).value

            # Skip placeholder row and empty rows
            if receipt_id is None and payment is None:
                continue
            if isinstance(receipt_id, str) and receipt_id.startswith('\u2190'):
                continue

            if receipt_id is not None and payment is not None and amount is not None:
                valid_rows += 1
                if isinstance(payment, str) and payment.strip() in EXPECTED_PAYMENT_METHODS:
                    payment_methods_found.add(payment.strip())
                try:
                    amount_float = float(amount)
                    if amount_float > 0:
                        valid_amounts.append(amount_float)
                except (ValueError, TypeError):
                    pass

        if valid_rows >= 12:
            print(f"PASS: Component 1a — all 12 transaction rows populated (found {valid_rows})")
            total_score += 0.20
        elif valid_rows >= 10:
            print(f"PARTIAL: Component 1a — {valid_rows}/12 transaction rows populated")
            total_score += 0.12
        elif valid_rows >= 6:
            print(f"PARTIAL: Component 1a — {valid_rows}/12 transaction rows populated")
            total_score += 0.06
        else:
            print(f"FAIL: Component 1a — only {valid_rows}/12 rows have data")

    except Exception as e:
        print(f"ERROR: Component 1a (row count) — {e}")

    # -------------------------------------------------------------------------
    # Component 1b: Payment methods diversity (0.08 points)
    # All 3 methods (Cash, Card, Mobile) must appear in the transactions
    # -------------------------------------------------------------------------
    try:
        methods_count = len(payment_methods_found)
        if methods_count == 3 and payment_methods_found == EXPECTED_PAYMENT_METHODS:
            print(f"PASS: Component 1b — all 3 payment methods present: {payment_methods_found}")
            total_score += 0.08
        elif methods_count >= 2:
            print(f"PARTIAL: Component 1b — only {methods_count}/3 payment methods: {payment_methods_found}")
            total_score += 0.04
        else:
            print(f"FAIL: Component 1b — missing payment methods, found: {payment_methods_found}")

    except Exception as e:
        print(f"ERROR: Component 1b (payment methods) — {e}")

    # -------------------------------------------------------------------------
    # Component 1c: Amount values are valid (0.07 points)
    # -------------------------------------------------------------------------
    try:
        if len(valid_amounts) >= 12:
            print(f"PASS: Component 1c — all 12 amounts are valid (sum={sum(valid_amounts):.2f})")
            total_score += 0.07
        elif len(valid_amounts) >= 8:
            print(f"PARTIAL: Component 1c — {len(valid_amounts)}/12 amounts valid")
            total_score += 0.04
        else:
            print(f"FAIL: Component 1c — only {len(valid_amounts)}/12 amounts are valid numbers")

    except Exception as e:
        print(f"ERROR: Component 1c (amounts) — {e}")

    # -------------------------------------------------------------------------
    # Component 2: EOD Report SUMIF formulas by payment method (0.25 points)
    # EOD Report must contain:
    #   - SUMIF formula with "Cash" criteria referencing Transactions sheet
    #   - SUMIF formula with "Card" criteria referencing Transactions sheet
    #   - SUMIF formula with "Mobile" criteria referencing Transactions sheet
    #   - Grand Total SUM(B...:B...) formula summing the SUMIF results
    # -------------------------------------------------------------------------
    try:
        formulas = scan_eod_formulas(ws_eod)
        sumif_count = sum([
            1 if formulas['sumif_cash'] >= 1 else 0,
            1 if formulas['sumif_card'] >= 1 else 0,
            1 if formulas['sumif_mobile'] >= 1 else 0,
        ])

        if sumif_count == 3:
            print(f"PASS: Component 2a — all 3 SUMIF formulas present (Cash/Card/Mobile)")
            total_score += 0.18
        elif sumif_count == 2:
            print(f"PARTIAL: Component 2a — {sumif_count}/3 SUMIF formulas present")
            total_score += 0.10
        elif sumif_count == 1:
            print(f"PARTIAL: Component 2a — only {sumif_count}/3 SUMIF formulas present")
            total_score += 0.05
        else:
            print(f"FAIL: Component 2a — no SUMIF formulas for Cash/Card/Mobile found")

        if formulas['grand_total'] >= 1:
            print(f"PASS: Component 2b — Grand Total SUM formula present in EOD Report")
            total_score += 0.07
        else:
            print(f"FAIL: Component 2b — Grand Total SUM formula not found in EOD Report")

    except Exception as e:
        print(f"ERROR: Component 2 (SUMIF formulas) — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Summary statistics formulas (0.15 points)
    # EOD Report must contain:
    #   - COUNTA or COUNT formula referencing Transactions sheet (total transactions)
    #   - AVERAGE formula referencing Transactions sheet (avg transaction value)
    # -------------------------------------------------------------------------
    try:
        has_count = formulas['count_stat'] >= 1
        has_average = formulas['average_stat'] >= 1

        if has_count and has_average:
            print(f"PASS: Component 3 — both COUNT and AVERAGE formulas present in EOD Report")
            total_score += 0.15
        elif has_count or has_average:
            present_formula = 'COUNT' if has_count else 'AVERAGE'
            print(f"PARTIAL: Component 3 — only {present_formula} formula present (missing the other)")
            total_score += 0.07
        else:
            print(f"FAIL: Component 3 — COUNT/AVERAGE formulas missing from EOD Report")

    except Exception as e:
        print(f"ERROR: Component 3 (summary statistics) — {e}")

    # -------------------------------------------------------------------------
    # Component 4: COUNTIFS for hourly transaction frequency (0.15 points)
    # EOD Report must contain 9 COUNTIFS formulas (one per hour slot 09:00-17:00)
    # each referencing Transactions sheet time column
    # -------------------------------------------------------------------------
    try:
        countifs_num = formulas['countifs']

        if countifs_num >= 9:
            print(f"PASS: Component 4 — all 9 COUNTIFS hourly formulas present (found {countifs_num})")
            total_score += 0.15
        elif countifs_num >= 6:
            print(f"PARTIAL: Component 4 — {countifs_num}/9 COUNTIFS formulas present")
            total_score += 0.09
        elif countifs_num >= 3:
            print(f"PARTIAL: Component 4 — {countifs_num}/9 COUNTIFS formulas present")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4 — only {countifs_num}/9 COUNTIFS hourly formulas present")

    except Exception as e:
        print(f"ERROR: Component 4 (COUNTIFS hourly) — {e}")

    # -------------------------------------------------------------------------
    # Component 5: Column chart of hourly transaction volume in EOD Report (0.10 points)
    # The EOD Report sheet must contain a BarChart with type="col" (vertical columns)
    # -------------------------------------------------------------------------
    try:
        charts = ws_eod._charts
        chart_count = len(charts)

        if chart_count >= 1:
            chart = charts[0]
            chart_class = type(chart).__name__

            # Determine if it's a vertical column chart (barDir='col' or type='col')
            bar_dir = getattr(chart, 'barDir', None)
            bar_type = getattr(chart, 'type', None)

            if bar_dir == 'col' or bar_type == 'col':
                print(f"PASS: Component 5 — column chart (barDir=col) present in EOD Report")
                total_score += 0.10
            elif chart_class in ('BarChart',):
                # BarChart present but direction not confirmed — still partial credit
                print(f"PARTIAL: Component 5 — BarChart present (barDir={bar_dir}) in EOD Report")
                total_score += 0.06
            elif chart_count >= 1:
                # Any chart type present — minimal partial credit
                print(f"PARTIAL: Component 5 — chart present but unexpected type ({chart_class})")
                total_score += 0.04
        else:
            print(f"FAIL: Component 5 — no chart found in EOD Report sheet")

    except Exception as e:
        print(f"ERROR: Component 5 (chart) — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score:.4f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point: test against canonical artifact path in the VM env
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
