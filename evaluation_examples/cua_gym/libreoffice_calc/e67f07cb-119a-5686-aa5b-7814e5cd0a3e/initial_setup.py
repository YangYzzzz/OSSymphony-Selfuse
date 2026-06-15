"""
Initial Setup: Read sales receipts and enter into daily_sales.xlsx
Task ID: osworld_multi_apps_receipt_to_calc_015
Domain: libreoffice_calc (multi-app: also creates receipt images/PDFs on desktop)

Creates:
  - /home/user/daily_sales.xlsx with empty Transactions sheet + pre-structured EOD Report sheet
  - 12 sales receipt files on Desktop (mix of JPEG and PDF)
  - Opens daily_sales.xlsx in LibreOffice Calc
"""

import os
import shlex
import subprocess
import time
import textwrap
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_receipt_to_calc_015'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'
DESKTOP = f'{WORKDIR}/Desktop'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def install_deps():
    """Install required Python packages."""
    subprocess.run(
        ['pip3', 'install', 'openpyxl', 'Pillow', 'fpdf2', '--quiet'],
        capture_output=True
    )


def create_receipt_jpeg(path, receipt_data):
    """Create a JPEG sales receipt image using Pillow."""
    from PIL import Image, ImageDraw, ImageFont

    width, height = 400, 280
    img = Image.new('RGB', (width, height), color=(255, 255, 255))
    draw = ImageDraw.Draw(img)

    # Try to get a font; fall back to default
    try:
        font_title = ImageFont.truetype('/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf', 18)
        font_body = ImageFont.truetype('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf', 14)
    except Exception:
        font_title = ImageFont.load_default()
        font_body = ImageFont.load_default()

    # Draw receipt border
    draw.rectangle([10, 10, width - 10, height - 10], outline=(0, 0, 0), width=2)

    # Title
    draw.text((width // 2, 30), 'THE DAILY GRIND CAFE', fill=(0, 0, 0),
              font=font_title, anchor='mm')
    draw.text((width // 2, 52), '123 Main Street | Tel: 555-0142', fill=(80, 80, 80),
              font=font_body, anchor='mm')
    draw.line([(20, 65), (width - 20, 65)], fill=(0, 0, 0), width=1)

    # Receipt info
    y = 80
    draw.text((30, y), f"Date: {receipt_data['date']}", fill=(0, 0, 0), font=font_body)
    draw.text((30, y + 22), f"Time: {receipt_data['time']}", fill=(0, 0, 0), font=font_body)
    draw.text((30, y + 44), f"Receipt #: {receipt_data['receipt_no']}", fill=(0, 0, 0), font=font_body)

    draw.line([(20, y + 68), (width - 20, y + 68)], fill=(0, 0, 0), width=1)

    # Items
    y2 = y + 80
    draw.text((30, y2), receipt_data['item'], fill=(0, 0, 0), font=font_body)
    draw.text((width - 30, y2), f"${receipt_data['amount']:.2f}", fill=(0, 0, 0),
              font=font_body, anchor='ra')

    draw.line([(20, y2 + 25), (width - 20, y2 + 25)], fill=(0, 0, 0), width=1)

    # Total
    y3 = y2 + 35
    draw.text((30, y3), 'TOTAL:', fill=(0, 0, 0), font=font_title)
    draw.text((width - 30, y3), f"${receipt_data['amount']:.2f}", fill=(0, 0, 0),
              font=font_title, anchor='ra')

    # Payment method
    y4 = y3 + 30
    draw.text((30, y4), f"Payment: {receipt_data['payment']}", fill=(0, 0, 128), font=font_body)

    # Thank you
    draw.text((width // 2, height - 25), 'Thank you for your visit!', fill=(80, 80, 80),
              font=font_body, anchor='mm')

    img.save(path, 'JPEG', quality=85)


def create_receipt_pdf(path, receipt_data):
    """Create a PDF sales receipt using fpdf2."""
    from fpdf import FPDF

    pdf = FPDF()
    pdf.add_page()
    pdf.set_margins(20, 20, 20)

    # Title
    pdf.set_font('Helvetica', 'B', 16)
    pdf.cell(0, 10, 'THE DAILY GRIND CAFE', align='C', ln=True)
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(0, 6, '123 Main Street | Tel: 555-0142', align='C', ln=True)
    pdf.ln(3)
    pdf.set_draw_color(0, 0, 0)
    pdf.line(20, pdf.get_y(), 190, pdf.get_y())
    pdf.ln(4)

    # Receipt details
    pdf.set_font('Helvetica', '', 11)
    pdf.cell(0, 7, f"Date: {receipt_data['date']}", ln=True)
    pdf.cell(0, 7, f"Time: {receipt_data['time']}", ln=True)
    pdf.cell(0, 7, f"Receipt #: {receipt_data['receipt_no']}", ln=True)
    pdf.ln(3)
    pdf.line(20, pdf.get_y(), 190, pdf.get_y())
    pdf.ln(4)

    # Item and amount
    pdf.set_font('Helvetica', '', 11)
    item_text = receipt_data['item']
    amount_text = f"${receipt_data['amount']:.2f}"
    pdf.cell(120, 8, item_text)
    pdf.cell(0, 8, amount_text, align='R', ln=True)

    pdf.ln(3)
    pdf.line(20, pdf.get_y(), 190, pdf.get_y())
    pdf.ln(4)

    # Total
    pdf.set_font('Helvetica', 'B', 13)
    pdf.cell(120, 8, 'TOTAL:')
    pdf.cell(0, 8, f"${receipt_data['amount']:.2f}", align='R', ln=True)
    pdf.ln(3)

    # Payment method
    pdf.set_font('Helvetica', 'I', 11)
    pdf.cell(0, 7, f"Payment Method: {receipt_data['payment']}", ln=True)

    pdf.ln(8)
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(0, 6, 'Thank you for your visit!', align='C', ln=True)

    pdf.output(path)


def create_receipts():
    """Create 12 sales receipt files on Desktop (6 JPEG + 6 PDF)."""
    os.makedirs(DESKTOP, exist_ok=True)

    # 12 sales transactions covering a business day
    # Grouped by payment type: 4 Cash, 4 Card, 4 Mobile
    # Times range from 09:xx to 17:xx
    receipts = [
        # JPEG receipts (6)
        {
            'receipt_no': 'R-001', 'date': '2025-03-15', 'time': '09:12',
            'item': 'Espresso x2 + Croissant', 'amount': 12.50, 'payment': 'Cash',
            'format': 'jpeg'
        },
        {
            'receipt_no': 'R-002', 'date': '2025-03-15', 'time': '09:47',
            'item': 'Latte + Blueberry Muffin', 'amount': 15.75, 'payment': 'Card',
            'format': 'jpeg'
        },
        {
            'receipt_no': 'R-003', 'date': '2025-03-15', 'time': '10:23',
            'item': 'Cappuccino x3 + Sandwich', 'amount': 31.20, 'payment': 'Mobile',
            'format': 'jpeg'
        },
        {
            'receipt_no': 'R-004', 'date': '2025-03-15', 'time': '11:05',
            'item': 'Flat White + Avocado Toast', 'amount': 22.40, 'payment': 'Cash',
            'format': 'jpeg'
        },
        {
            'receipt_no': 'R-005', 'date': '2025-03-15', 'time': '11:38',
            'item': 'Cold Brew + Cake Slice', 'amount': 18.90, 'payment': 'Card',
            'format': 'jpeg'
        },
        {
            'receipt_no': 'R-006', 'date': '2025-03-15', 'time': '12:15',
            'item': 'Iced Mocha x2 + Cookie', 'amount': 27.60, 'payment': 'Mobile',
            'format': 'jpeg'
        },
        # PDF receipts (6)
        {
            'receipt_no': 'R-007', 'date': '2025-03-15', 'time': '13:02',
            'item': 'Americano + Quiche', 'amount': 16.30, 'payment': 'Cash',
            'format': 'pdf'
        },
        {
            'receipt_no': 'R-008', 'date': '2025-03-15', 'time': '13:44',
            'item': 'Green Tea Latte x2 + Brownie', 'amount': 24.80, 'payment': 'Card',
            'format': 'pdf'
        },
        {
            'receipt_no': 'R-009', 'date': '2025-03-15', 'time': '14:20',
            'item': 'Macchiato + Banana Bread', 'amount': 14.50, 'payment': 'Mobile',
            'format': 'pdf'
        },
        {
            'receipt_no': 'R-010', 'date': '2025-03-15', 'time': '15:08',
            'item': 'Chai Latte x2 + Scone', 'amount': 21.00, 'payment': 'Cash',
            'format': 'pdf'
        },
        {
            'receipt_no': 'R-011', 'date': '2025-03-15', 'time': '16:33',
            'item': 'Mocha Frappuccino x3', 'amount': 33.75, 'payment': 'Card',
            'format': 'pdf'
        },
        {
            'receipt_no': 'R-012', 'date': '2025-03-15', 'time': '17:10',
            'item': 'Hot Chocolate + Danish Pastry x2', 'amount': 19.45, 'payment': 'Mobile',
            'format': 'pdf'
        },
    ]

    for r in receipts:
        receipt_no = r['receipt_no']
        if r['format'] == 'jpeg':
            path = os.path.join(DESKTOP, f"receipt_{receipt_no}.jpg")
            create_receipt_jpeg(path, r)
            print(f"Created JPEG receipt: {path}")
        else:
            path = os.path.join(DESKTOP, f"receipt_{receipt_no}.pdf")
            create_receipt_pdf(path, r)
            print(f"Created PDF receipt: {path}")


def create_initial_spreadsheet():
    """Create daily_sales.xlsx with empty Transactions sheet and pre-structured EOD Report."""
    wb = openpyxl.Workbook()

    # -------------------------------------------------------
    # Sheet 1: Transactions
    # -------------------------------------------------------
    ws_trans = wb.active
    ws_trans.title = 'Transactions'

    # Header styling
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF1F4E79', end_color='FF1F4E79', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers_trans = ['Receipt #', 'Date', 'Time', 'Payment Method', 'Item(s)', 'Amount ($)']
    col_widths = [12, 14, 10, 18, 40, 14]

    for col, (h, w) in enumerate(zip(headers_trans, col_widths), 1):
        cell = ws_trans.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border
        ws_trans.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    ws_trans.row_dimensions[1].height = 30
    ws_trans.freeze_panes = 'A2'

    # Instructions text in row 2 (light gray, to be replaced by agent with data)
    instr_font = Font(name='Calibri', size=10, italic=True, color='888888')
    cell = ws_trans.cell(row=2, column=1, value='← Enter receipt data here (Rows 2-13)')
    cell.font = instr_font

    # -------------------------------------------------------
    # Sheet 2: EOD Report (pre-structured, no formulas yet)
    # -------------------------------------------------------
    ws_eod = wb.create_sheet('EOD Report')

    # Title
    ws_eod.merge_cells('A1:D1')
    title_cell = ws_eod['A1']
    title_cell.value = 'THE DAILY GRIND CAFE — End of Day Report'
    title_cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='FF1F4E79', end_color='FF1F4E79', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_eod.row_dimensions[1].height = 32

    # Date sub-header
    ws_eod.merge_cells('A2:D2')
    ws_eod['A2'].value = 'Date: 2025-03-15'
    ws_eod['A2'].font = Font(name='Calibri', size=11, italic=True)
    ws_eod['A2'].alignment = Alignment(horizontal='center')

    # Section: Sales by Payment Method
    ws_eod['A4'].value = 'SALES BY PAYMENT METHOD'
    ws_eod['A4'].font = Font(name='Calibri', size=11, bold=True)
    ws_eod['A4'].fill = PatternFill(start_color='FFD6E4F0', end_color='FFD6E4F0', fill_type='solid')
    ws_eod.merge_cells('A4:D4')

    # Payment method rows — labels only, totals left blank for agent
    section_font = Font(name='Calibri', size=11)
    label_fill = PatternFill(start_color='FFEFF7FF', end_color='FFEFF7FF', fill_type='solid')

    ws_eod['A5'].value = 'Cash Sales Total'
    ws_eod['A6'].value = 'Card Sales Total'
    ws_eod['A7'].value = 'Mobile Sales Total'
    ws_eod['A8'].value = 'Grand Total'

    for row in range(5, 9):
        ws_eod.cell(row=row, column=1).font = section_font
        ws_eod.cell(row=row, column=1).fill = label_fill
        # Column B: formula placeholder (left blank — agent fills)
        ws_eod.cell(row=row, column=2).number_format = '$#,##0.00'

    ws_eod['A8'].font = Font(name='Calibri', size=11, bold=True)

    # Section: Summary Statistics
    ws_eod['A10'].value = 'SUMMARY STATISTICS'
    ws_eod['A10'].font = Font(name='Calibri', size=11, bold=True)
    ws_eod['A10'].fill = PatternFill(start_color='FFD6E4F0', end_color='FFD6E4F0', fill_type='solid')
    ws_eod.merge_cells('A10:D10')

    ws_eod['A11'].value = 'Total Transactions'
    ws_eod['A12'].value = 'Average Transaction Value'
    ws_eod['A13'].value = 'Highest Single Transaction'
    ws_eod['A14'].value = 'Lowest Single Transaction'

    for row in range(11, 15):
        ws_eod.cell(row=row, column=1).font = section_font
        ws_eod.cell(row=row, column=1).fill = label_fill
        ws_eod.cell(row=row, column=2).number_format = '$#,##0.00'

    # Section: Hourly Transaction Count (agent uses COUNTIFS)
    ws_eod['A16'].value = 'HOURLY TRANSACTION FREQUENCY'
    ws_eod['A16'].font = Font(name='Calibri', size=11, bold=True)
    ws_eod['A16'].fill = PatternFill(start_color='FFD6E4F0', end_color='FFD6E4F0', fill_type='solid')
    ws_eod.merge_cells('A16:D16')

    ws_eod['A17'].value = 'Hour'
    ws_eod['B17'].value = 'Transaction Count'
    ws_eod['A17'].font = Font(name='Calibri', size=11, bold=True)
    ws_eod['B17'].font = Font(name='Calibri', size=11, bold=True)

    # Hour slots for business day (9AM - 5PM) — counts left blank for agent
    hours = ['09:00-10:00', '10:00-11:00', '11:00-12:00', '12:00-13:00',
             '13:00-14:00', '14:00-15:00', '15:00-16:00', '16:00-17:00', '17:00-18:00']
    for i, hour in enumerate(hours, 18):
        ws_eod.cell(row=i, column=1, value=hour).font = section_font
        ws_eod.cell(row=i, column=1).fill = label_fill

    # Placeholder note for chart area
    ws_eod['D18'].value = '← Chart will be placed here'
    ws_eod['D18'].font = Font(name='Calibri', size=10, italic=True, color='888888')

    # Column widths
    ws_eod.column_dimensions['A'].width = 28
    ws_eod.column_dimensions['B'].width = 22
    ws_eod.column_dimensions['C'].width = 16
    ws_eod.column_dimensions['D'].width = 30

    # -------------------------------------------------------
    # Save workbook
    # -------------------------------------------------------
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


def create_initial():
    install_deps()

    # Create the spreadsheet
    create_initial_spreadsheet()

    # Create receipt files on Desktop
    create_receipts()

    # GUI-ready startup: open daily_sales.xlsx in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=3.0)

    # Also open the Desktop folder in file manager so receipts are visible
    launch_gui(f'nautilus "{DESKTOP}"', delay_sec=1.5)

    print('GUI_READY: launched LibreOffice Calc with daily_sales.xlsx and Nautilus showing Desktop receipts')


create_initial()
