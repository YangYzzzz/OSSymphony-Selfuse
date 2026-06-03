"""
Initial Setup: Create invoice spreadsheet with data in A1:E20
Task ID: calc_gao_048
Domain: libreoffice_calc

Creates an invoice summary spreadsheet. Cell E10 contains $5,432.10 (tax-inclusive).
No shapes or callouts exist. Area around G10 is empty.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_gao_048'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # --- Styles ---
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    currency_fmt = '$#,##0.00'
    date_fmt = 'yyyy-mm-dd'
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # --- Headers (Row 1) ---
    headers = ["Invoice #", "Description", "Quantity", "Unit Price", "Total"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Invoice Data (Rows 2-20) ---
    # Realistic invoice line items for a consulting/services company
    data = [
        ["INV-2025-001", "Website Redesign - Phase 1", 1, 3500.00, 3500.00],
        ["INV-2025-002", "SEO Audit & Optimization", 1, 1200.00, 1200.00],
        ["INV-2025-003", "Monthly Hosting (Q1)", 3, 89.99, 269.97],
        ["INV-2025-004", "Logo Design Package", 1, 850.00, 850.00],
        ["INV-2025-005", "Content Writing - Blog Posts", 8, 125.00, 1000.00],
        ["INV-2025-006", "Email Campaign Setup", 1, 475.00, 475.00],
        ["INV-2025-007", "Social Media Management (Mar)", 1, 1800.00, 1800.00],
        ["INV-2025-008", "Photography Session", 1, 650.00, 650.00],
        ["INV-2025-009", "Annual Software License (incl. tax)", 1, 5432.10, 5432.10],  # Row 10, E10 = $5,432.10
        ["INV-2025-010", "Print Materials - Brochures", 500, 1.25, 625.00],
        ["INV-2025-011", "Video Production - 30s Ad", 1, 2200.00, 2200.00],
        ["INV-2025-012", "CRM Integration Service", 1, 1500.00, 1500.00],
        ["INV-2025-013", "Staff Training Workshop", 2, 400.00, 800.00],
        ["INV-2025-014", "Domain Registration (5 yr)", 1, 74.95, 74.95],
        ["INV-2025-015", "SSL Certificate Renewal", 3, 49.99, 149.97],
        ["INV-2025-016", "Analytics Dashboard Setup", 1, 950.00, 950.00],
        ["INV-2025-017", "Mobile App Consultation", 4, 175.00, 700.00],
        ["INV-2025-018", "Database Migration", 1, 1350.00, 1350.00],
        ["INV-2025-019", "Quarterly Performance Report", 1, 600.00, 600.00],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 1:  # Invoice #
                cell.alignment = Alignment(horizontal="center")
            elif c == 3:  # Quantity
                cell.alignment = Alignment(horizontal="center")
            elif c in (4, 5):  # Unit Price, Total
                cell.number_format = currency_fmt

    # --- Column Widths ---
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    # --- Freeze top row ---
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Verify E10
    wb_check = openpyxl.load_workbook(OUTPUT)
    val_e10 = wb_check["Invoice"]["E10"].value
    print(f'E10 value: {val_e10} (expected 5432.10)')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
