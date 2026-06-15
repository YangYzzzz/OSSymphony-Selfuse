"""
Reward Script: Create pivot table from research data showing count and average
              of measurements per experiment group per treatment type.
Task ID: calc_pivot_050
Domain: libreoffice_calc
Scoring:
  Component 1: PivotTable sheet exists (0.15)
  Component 2: Row structure - ExperimentGroup labels (0.15)
  Component 3: Column structure - Treatment headers for COUNT and AVERAGE (0.15)
  Component 4: COUNT values correct (0.25)
  Component 5: AVERAGE values correct (0.20)
  Component 6: Grand Total row correct (0.10)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_050'

# Expected groups and treatments
EXPECTED_GROUPS = ['Control', 'GroupA', 'GroupB', 'GroupC']
EXPECTED_TREATMENTS = ['Placebo', 'Low', 'Medium', 'High']

# Expected COUNT values (each group has 15 per treatment, 60 total; grand total 240)
EXPECTED_COUNTS = {
    'Control':     {'Placebo': 15, 'Low': 15, 'Medium': 15, 'High': 15, 'Total': 60},
    'GroupA':      {'Placebo': 15, 'Low': 15, 'Medium': 15, 'High': 15, 'Total': 60},
    'GroupB':      {'Placebo': 15, 'Low': 15, 'Medium': 15, 'High': 15, 'Total': 60},
    'GroupC':      {'Placebo': 15, 'Low': 15, 'Medium': 15, 'High': 15, 'Total': 60},
}
GRAND_TOTAL_COUNT = 240

# Expected AVERAGE values (from golden with tolerance)
EXPECTED_AVERAGES = {
    'Control':     {'Placebo': 3.2, 'Low': 4.1, 'Medium': 5.5, 'High': 7.2, 'Total': 5.0},
    'GroupA':      {'Placebo': 3.5, 'Low': 5.8, 'Medium': 7.3, 'High': 9.1, 'Total': 6.42},
    'GroupB':      {'Placebo': 3.1, 'Low': 6.2, 'Medium': 8.0, 'High': 10.5, 'Total': 6.95},
    'GroupC':      {'Placebo': 2.9, 'Low': 5.0, 'Medium': 7.8, 'High': 11.2, 'Total': 6.72},
}
GRAND_TOTAL_AVG = 6.28
AVG_TOLERANCE = 0.5  # tolerance for average values


def find_pivot_sheet(wb):
    """Find a sheet that looks like a pivot table (not ExperimentData)."""
    for sn in wb.sheetnames:
        if sn.lower() != 'experimentdata':
            ws = wb[sn]
            # Check if it has some data that looks like a pivot table
            if ws.max_row >= 3 and ws.max_column >= 3:
                return ws
    return None


def scan_for_groups_and_data(ws):
    """
    Scan the pivot sheet to locate group labels, treatment headers,
    COUNT section, and AVERAGE section dynamically.
    Returns a dict with parsed structure or None if not found.
    """
    result = {
        'groups_found': [],
        'group_rows': {},
        'count_cols': {},
        'avg_cols': {},
        'header_row': None,
        'grand_total_row': None,
    }

    # Scan column A for group names and grand total
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val is None:
            continue
        val_str = str(val).strip()
        for grp in EXPECTED_GROUPS:
            if val_str.lower() == grp.lower():
                result['groups_found'].append(grp)
                result['group_rows'][grp] = r
        if 'grand' in val_str.lower() and 'total' in val_str.lower():
            result['grand_total_row'] = r

    # Find treatment header row (row with Placebo, Low, Medium, High)
    for r in range(1, ws.max_row + 1):
        row_vals = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                row_vals.append((c, str(v).strip()))
        treatments_found = [col for col, v in row_vals if v in EXPECTED_TREATMENTS]
        if len(treatments_found) >= 4:
            result['header_row'] = r
            # Map treatment names to columns
            # There should be two sets: one for COUNT, one for AVERAGE
            treatment_positions = []
            for c, v in row_vals:
                if v in EXPECTED_TREATMENTS or v.lower() == 'total':
                    treatment_positions.append((c, v))

            # Split into two groups (COUNT section and AVERAGE section)
            if len(treatment_positions) >= 8:
                # First 5 are COUNT (4 treatments + total), next 5 are AVERAGE
                count_section = treatment_positions[:5]
                avg_section = treatment_positions[5:]
                for c, v in count_section:
                    result['count_cols'][v] = c
                for c, v in avg_section:
                    result['avg_cols'][v] = c
            elif len(treatment_positions) >= 4:
                # Maybe just one section found, try to use row above for labels
                for c, v in treatment_positions[:5]:
                    result['count_cols'][v] = c
            break

    return result


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: PivotTable sheet exists (0.15 points)
    # Initial file has only 'ExperimentData'. Golden has an additional pivot sheet.
    try:
        pivot_ws = find_pivot_sheet(wb)
        if pivot_ws is not None:
            print(f"PASS: Component 1 — Pivot sheet found: '{pivot_ws.title}' (0.15 pts)")
            total_score += 0.15
        else:
            print("FAIL: Component 1 — No pivot table sheet found (only ExperimentData)")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print(f"REWARD: {total_score}")
        return total_score

    # Scan the pivot sheet structure
    structure = scan_for_groups_and_data(pivot_ws)

    # Component 2: Row structure - ExperimentGroup labels present (0.15 points)
    try:
        found_groups = structure['groups_found']
        matched = [g for g in EXPECTED_GROUPS if g in found_groups]
        if len(matched) == 4:
            print(f"PASS: Component 2 — All 4 experiment groups found in rows (0.15 pts)")
            total_score += 0.15
        elif len(matched) >= 2:
            partial = 0.15 * len(matched) / 4
            print(f"PARTIAL: Component 2 — {len(matched)}/4 groups found: {matched} ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {len(matched)} groups found: {matched}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Column structure - Treatment headers for COUNT and AVERAGE (0.15 points)
    try:
        count_treatments = [t for t in EXPECTED_TREATMENTS if t in structure['count_cols']]
        avg_treatments = [t for t in EXPECTED_TREATMENTS if t in structure['avg_cols']]

        count_ok = len(count_treatments) >= 4
        avg_ok = len(avg_treatments) >= 4

        if count_ok and avg_ok:
            print(f"PASS: Component 3 — Both COUNT and AVERAGE sections have treatment headers (0.15 pts)")
            total_score += 0.15
        elif count_ok or avg_ok:
            print(f"PARTIAL: Component 3 — COUNT headers: {count_ok}, AVERAGE headers: {avg_ok} (0.075 pts)")
            total_score += 0.075
        else:
            print(f"FAIL: Component 3 — Treatment headers not properly structured. Count cols: {structure['count_cols']}, Avg cols: {structure['avg_cols']}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: COUNT values correct (0.25 points)
    try:
        if structure['count_cols'] and structure['group_rows']:
            count_correct = 0
            count_total = 0
            for grp in EXPECTED_GROUPS:
                if grp not in structure['group_rows']:
                    continue
                row = structure['group_rows'][grp]
                for treat in EXPECTED_TREATMENTS + ['Total']:
                    if treat not in structure['count_cols']:
                        continue
                    col = structure['count_cols'][treat]
                    cell_val = pivot_ws.cell(row=row, column=col).value
                    expected = EXPECTED_COUNTS.get(grp, {}).get(treat)
                    if expected is not None:
                        count_total += 1
                        if cell_val is not None:
                            try:
                                if abs(float(cell_val) - expected) < 0.5:
                                    count_correct += 1
                                else:
                                    print(f"  COUNT mismatch: {grp}/{treat} expected={expected}, got={cell_val}")
                            except (ValueError, TypeError):
                                print(f"  COUNT type error: {grp}/{treat} got={cell_val}")

            if count_total > 0:
                ratio = count_correct / count_total
                pts = 0.25 * ratio
                if ratio >= 0.9:
                    print(f"PASS: Component 4 — COUNT values: {count_correct}/{count_total} correct ({pts:.3f} pts)")
                    total_score += pts
                elif ratio > 0:
                    print(f"PARTIAL: Component 4 — COUNT values: {count_correct}/{count_total} correct ({pts:.3f} pts)")
                    total_score += pts
            else:
                print("FAIL: Component 4 — No COUNT values could be checked")
        else:
            print("FAIL: Component 4 — COUNT section not found in pivot table")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: AVERAGE values correct (0.20 points)
    try:
        if structure['avg_cols'] and structure['group_rows']:
            avg_correct = 0
            avg_total = 0
            for grp in EXPECTED_GROUPS:
                if grp not in structure['group_rows']:
                    continue
                row = structure['group_rows'][grp]
                for treat in EXPECTED_TREATMENTS + ['Total']:
                    if treat not in structure['avg_cols']:
                        continue
                    col = structure['avg_cols'][treat]
                    cell_val = pivot_ws.cell(row=row, column=col).value
                    expected = EXPECTED_AVERAGES.get(grp, {}).get(treat)
                    if expected is not None:
                        avg_total += 1
                        if cell_val is not None:
                            try:
                                if abs(float(cell_val) - expected) <= AVG_TOLERANCE:
                                    avg_correct += 1
                                else:
                                    print(f"  AVG mismatch: {grp}/{treat} expected={expected}, got={cell_val}")
                            except (ValueError, TypeError):
                                print(f"  AVG type error: {grp}/{treat} got={cell_val}")

            if avg_total > 0:
                ratio = avg_correct / avg_total
                pts = 0.20 * ratio
                if ratio >= 0.9:
                    print(f"PASS: Component 5 — AVERAGE values: {avg_correct}/{avg_total} correct ({pts:.3f} pts)")
                    total_score += pts
                elif ratio > 0:
                    print(f"PARTIAL: Component 5 — AVERAGE values: {avg_correct}/{avg_total} correct ({pts:.3f} pts)")
                    total_score += pts
            else:
                print("FAIL: Component 5 — No AVERAGE values could be checked")
        else:
            print("FAIL: Component 5 — AVERAGE section not found in pivot table")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Grand Total row correct (0.10 points)
    try:
        gt_row = structure['grand_total_row']
        if gt_row is not None:
            gt_pts = 0.0
            # Check grand total COUNT = 240
            if 'Total' in structure['count_cols']:
                gt_count_col = structure['count_cols']['Total']
                gt_count = pivot_ws.cell(row=gt_row, column=gt_count_col).value
                if gt_count is not None and abs(float(gt_count) - GRAND_TOTAL_COUNT) < 0.5:
                    gt_pts += 0.05
                    print(f"  Grand Total COUNT = {gt_count} (expected {GRAND_TOTAL_COUNT}) OK")
                else:
                    print(f"  Grand Total COUNT mismatch: expected {GRAND_TOTAL_COUNT}, got {gt_count}")

            # Check grand total AVERAGE ~ 6.28
            if 'Total' in structure['avg_cols']:
                gt_avg_col = structure['avg_cols']['Total']
                gt_avg = pivot_ws.cell(row=gt_row, column=gt_avg_col).value
                if gt_avg is not None and abs(float(gt_avg) - GRAND_TOTAL_AVG) <= AVG_TOLERANCE:
                    gt_pts += 0.05
                    print(f"  Grand Total AVG = {gt_avg} (expected ~{GRAND_TOTAL_AVG}) OK")
                else:
                    print(f"  Grand Total AVG mismatch: expected ~{GRAND_TOTAL_AVG}, got {gt_avg}")

            if gt_pts > 0:
                print(f"PASS: Component 6 — Grand Total row ({gt_pts:.2f} pts)")
                total_score += gt_pts
            else:
                print("FAIL: Component 6 — Grand Total values incorrect")
        else:
            print("FAIL: Component 6 — Grand Total row not found")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score:.3f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
