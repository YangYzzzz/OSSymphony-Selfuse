"""
Initial Setup: Create experiment data spreadsheet for pivot table task
Task ID: calc_pivot_050
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_050'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)

    groups = ['Control', 'GroupA', 'GroupB', 'GroupC']
    treatments = ['Placebo', 'Low', 'Medium', 'High']
    technicians = [
        'Sarah Chen', 'Marcus Johnson', 'Elena Rodriguez',
        'David Kim', 'Priya Patel', 'James Wilson'
    ]

    # We need 240 rows total, evenly distributed: 4 groups x 4 treatments = 16 combos
    # 240 / 16 = 15 per combo
    rows_per_combo = 15

    # For Control/Placebo we need AVG=3.2 exactly
    # Generate 15 values that average to 3.2 (sum=48.0)
    def generate_values_with_avg(n, target_avg, low=0.5, high=15.0):
        target_sum = round(target_avg * n, 4)
        values = []
        for i in range(n - 1):
            remaining = n - len(values) - 1
            current_sum = sum(values)
            needed = target_sum - current_sum
            # Calculate bounds for this value
            min_val = max(low, needed - remaining * high)
            max_val = min(high, needed - remaining * low)
            val = round(random.uniform(min_val, max_val), 2)
            values.append(val)
        # Last value to hit exact sum
        last_val = round(target_sum - sum(values), 2)
        values.append(last_val)
        random.shuffle(values)
        return values

    # Generate measurements for each group/treatment combo
    # Define approximate means for each combo to make data realistic
    combo_avgs = {
        ('Control', 'Placebo'): 3.2,   # exact per ground truth
        ('Control', 'Low'): 4.1,
        ('Control', 'Medium'): 5.5,
        ('Control', 'High'): 7.2,
        ('GroupA', 'Placebo'): 3.5,
        ('GroupA', 'Low'): 5.8,
        ('GroupA', 'Medium'): 7.3,
        ('GroupA', 'High'): 9.1,
        ('GroupB', 'Placebo'): 3.1,
        ('GroupB', 'Low'): 6.2,
        ('GroupB', 'Medium'): 8.0,
        ('GroupB', 'High'): 10.5,
        ('GroupC', 'Placebo'): 2.9,
        ('GroupC', 'Low'): 5.0,
        ('GroupC', 'Medium'): 7.8,
        ('GroupC', 'High'): 11.2,
    }

    # Build all data rows
    all_rows = []
    sample_id = 1
    base_date = datetime(2025, 1, 6)

    for group in groups:
        for treatment in treatments:
            avg = combo_avgs[(group, treatment)]
            measurements = generate_values_with_avg(rows_per_combo, avg)
            for m in measurements:
                date = base_date + timedelta(days=random.randint(0, 180))
                tech = random.choice(technicians)
                all_rows.append([sample_id, group, treatment, m, date, tech])
                sample_id += 1

    # Shuffle rows so they look like real unsorted experiment data
    random.shuffle(all_rows)
    # Re-assign sequential SampleIDs after shuffle
    for i, row in enumerate(all_rows):
        row[0] = i + 1

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ExperimentData'

    # Headers
    headers = ['SampleID', 'ExperimentGroup', 'Treatment', 'Measurement', 'Date', 'Technician']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # Data rows
    for r, row_data in enumerate(all_rows, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 4:  # Measurement
                cell.number_format = '0.00'
            elif c == 5:  # Date
                cell.number_format = 'yyyy-mm-dd'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 18

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Total rows: {len(all_rows)}')

    # Verify Control/Placebo stats
    cp_vals = [r[3] for r in all_rows if r[1] == 'Control' and r[2] == 'Placebo']
    print(f'Control/Placebo count: {len(cp_vals)}, avg: {sum(cp_vals)/len(cp_vals):.4f}')
    print(f'Total count: {len(all_rows)}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
