"""
Reward Script: VLOOKUP Auto-fill Product Details (SKU Lookup)
Task ID: calc_sales_product_sku_lookup_020
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): B2:B101 - IFERROR(VLOOKUP(A*,ProductMaster.$A:$D,2,0),"SKU Not Found")
  Component 2 (0.25): C2:C101 - IFERROR(VLOOKUP(A*,ProductMaster.$A:$D,3,0),"SKU Not Found")
  Component 3 (0.25): D2:D101 - IFERROR(VLOOKUP(A*,ProductMaster.$A:$D,4,0),0)
  Component 4 (0.15): G2:G101 - =D*(1-F*) Net Price formulas
Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_product_sku_lookup_020'


def normalize_formula(formula):
    """Remove spaces and uppercase for comparison."""
    if formula is None:
        return ''
    return str(formula).strip().upper().replace(' ', '')


def check_vlookup_formula(formula, row, col_index, error_fallback):
    """
    Check if formula matches expected IFERROR(VLOOKUP(A<row>,ProductMaster.$A:$D,<col_index>,0),<error_fallback>).
    Accepts both exact match and functionally equivalent forms.
    """
    norm = normalize_formula(formula)
    # Expected: =IFERROR(VLOOKUP(A<row>,ProductMaster.$A:$D,<col_index>,0),<error_fallback>)
    expected = normalize_formula(
        f'=IFERROR(VLOOKUP(A{row},ProductMaster.$A:$D,{col_index},0),{error_fallback})'
    )
    if norm == expected:
        return True

    # Also accept with FALSE instead of 0 for exact match param
    expected_false = normalize_formula(
        f'=IFERROR(VLOOKUP(A{row},ProductMaster.$A:$D,{col_index},FALSE),{error_fallback})'
    )
    if norm == expected_false:
        return True

    # Also accept ProductMaster!$A:$D (without dot notation)
    expected_excl = normalize_formula(
        f'=IFERROR(VLOOKUP(A{row},PRODUCTMASTER!$A:$D,{col_index},0),{error_fallback})'
    )
    if norm == expected_excl:
        return True

    expected_excl_false = normalize_formula(
        f'=IFERROR(VLOOKUP(A{row},PRODUCTMASTER!$A:$D,{col_index},FALSE),{error_fallback})'
    )
    if norm == expected_excl_false:
        return True

    # Accept also with more specific range like $A$1:$D$76
    # Check if it has the right structure with regex
    # Must start with =IFERROR(VLOOKUP(A<row>,...,col_index,...)
    pattern_dot = re.compile(
        r'^=IFERROR\(VLOOKUP\(A' + str(row) + r',PRODUCTMASTER\.\$A[:\$].*\$D[^,]*,' + str(col_index) + r',(0|FALSE)\)',
        re.IGNORECASE
    )
    pattern_excl = re.compile(
        r'^=IFERROR\(VLOOKUP\(A' + str(row) + r',PRODUCTMASTER!\$A[:\$].*\$D[^,]*,' + str(col_index) + r',(0|FALSE)\)',
        re.IGNORECASE
    )
    stripped = str(formula).strip()
    if pattern_dot.match(stripped) or pattern_excl.match(stripped):
        return True

    return False


def check_net_price_formula(formula, row):
    """
    Check if G<row> has =D<row>*(1-F<row>) or equivalent.
    """
    norm = normalize_formula(formula)
    expected = normalize_formula(f'=D{row}*(1-F{row})')
    if norm == expected:
        return True
    # Also accept =D<row>*(1-F<row>) with spaces
    return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition gate: file must load
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: required sheets must exist
    if 'Orders' not in wb.sheetnames:
        print("CRITICAL: 'Orders' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    if 'ProductMaster' not in wb.sheetnames:
        print("CRITICAL: 'ProductMaster' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws_orders = wb['Orders']

    # -----------------------------------------------------------------------
    # Component 1: B2:B101 — IFERROR(VLOOKUP(A*,ProductMaster.$A:$D,2,0),"SKU Not Found")
    #   Full 100/100 = 0.35 pts; partial credit for partial coverage
    # -----------------------------------------------------------------------
    try:
        b_correct = 0
        b_total = 100
        b_failures = []

        for row in range(2, 102):
            formula = ws_orders.cell(row=row, column=2).value
            if check_vlookup_formula(formula, row, 2, '"SKU Not Found"'):
                b_correct += 1
            else:
                if len(b_failures) < 3:
                    b_failures.append(f"B{row}: {repr(formula)}")

        if b_correct == b_total:
            print(f"PASS: Component 1 — B2:B101 all have VLOOKUP col-2 with error handling ({b_correct}/100 rows) (0.35 pts)")
            total_score += 0.35
        elif b_correct >= 90:
            # Partial credit for 90%+ coverage
            partial = round(0.35 * b_correct / b_total, 4)
            print(f"PARTIAL: Component 1 — B2:B101 {b_correct}/100 rows correct. Partial: {partial} pts")
            print(f"  Sample failures: {b_failures}")
            if partial > 0:
                total_score += partial
        else:
            print(f"FAIL: Component 1 — B2:B101 only {b_correct}/100 rows correct")
            print(f"  Sample failures: {b_failures}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: C2:C101 — IFERROR(VLOOKUP(A*,ProductMaster.$A:$D,3,0),"SKU Not Found")
    # -----------------------------------------------------------------------
    try:
        c_correct = 0
        c_total = 100
        c_failures = []

        for row in range(2, 102):
            formula = ws_orders.cell(row=row, column=3).value
            if check_vlookup_formula(formula, row, 3, '"SKU Not Found"'):
                c_correct += 1
            else:
                if len(c_failures) < 3:
                    c_failures.append(f"C{row}: {repr(formula)}")

        if c_correct == c_total:
            print(f"PASS: Component 2 — C2:C101 all have VLOOKUP col-3 with error handling ({c_correct}/100 rows) (0.25 pts)")
            total_score += 0.25
        elif c_correct >= 90:
            partial = round(0.25 * c_correct / c_total, 4)
            print(f"PARTIAL: Component 2 — C2:C101 {c_correct}/100 rows correct. Partial: {partial} pts")
            print(f"  Sample failures: {c_failures}")
            if partial > 0:
                total_score += partial
        else:
            print(f"FAIL: Component 2 — C2:C101 only {c_correct}/100 rows correct")
            print(f"  Sample failures: {c_failures}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: D2:D101 — IFERROR(VLOOKUP(A*,ProductMaster.$A:$D,4,0),0)
    #   Note: fallback for List Price is 0 (not "SKU Not Found")
    # -----------------------------------------------------------------------
    try:
        d_correct = 0
        d_total = 100
        d_failures = []

        for row in range(2, 102):
            formula = ws_orders.cell(row=row, column=4).value
            if check_vlookup_formula(formula, row, 4, '0'):
                d_correct += 1
            else:
                if len(d_failures) < 3:
                    d_failures.append(f"D{row}: {repr(formula)}")

        if d_correct == d_total:
            print(f"PASS: Component 3 — D2:D101 all have VLOOKUP col-4 with 0 fallback ({d_correct}/100 rows) (0.25 pts)")
            total_score += 0.25
        elif d_correct >= 90:
            partial = round(0.25 * d_correct / d_total, 4)
            print(f"PARTIAL: Component 3 — D2:D101 {d_correct}/100 rows correct. Partial: {partial} pts")
            print(f"  Sample failures: {d_failures}")
            if partial > 0:
                total_score += partial
        else:
            print(f"FAIL: Component 3 — D2:D101 only {d_correct}/100 rows correct")
            print(f"  Sample failures: {d_failures}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: G2:G101 — =D*(1-F*) Net Price formula
    # -----------------------------------------------------------------------
    try:
        g_correct = 0
        g_total = 100
        g_failures = []

        for row in range(2, 102):
            formula = ws_orders.cell(row=row, column=7).value
            if check_net_price_formula(formula, row):
                g_correct += 1
            else:
                if len(g_failures) < 3:
                    g_failures.append(f"G{row}: {repr(formula)}")

        if g_correct == g_total:
            print(f"PASS: Component 4 — G2:G101 all have Net Price formula =D*(1-F*) ({g_correct}/100 rows) (0.15 pts)")
            total_score += 0.15
        elif g_correct >= 90:
            partial = round(0.15 * g_correct / g_total, 4)
            print(f"PARTIAL: Component 4 — G2:G101 {g_correct}/100 rows correct. Partial: {partial} pts")
            print(f"  Sample failures: {g_failures}")
            if partial > 0:
                total_score += partial
        else:
            print(f"FAIL: Component 4 — G2:G101 only {g_correct}/100 rows correct")
            print(f"  Sample failures: {g_failures}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
