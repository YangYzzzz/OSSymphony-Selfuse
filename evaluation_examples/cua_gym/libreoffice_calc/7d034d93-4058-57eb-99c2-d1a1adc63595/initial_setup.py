"""
Initial Setup: Sales orders spreadsheet with product SKU lookup
Task ID: calc_sales_product_sku_lookup_020
Domain: libreoffice_calc

Creates:
- Sheet 'Orders': 100 order rows with SKUs in col A, empty B/C/D/G, data in E and F
- Sheet 'ProductMaster': 75 products with SKU, Product Name, Category, List Price, Min Order
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_product_sku_lookup_020'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # ── ProductMaster sheet (created first so SKUs are available) ──
    ws_pm = wb.active
    ws_pm.title = 'ProductMaster'

    pm_headers = ['SKU', 'Product Name', 'Category', 'List Price', 'Min Order']
    for col, h in enumerate(pm_headers, 1):
        cell = ws_pm.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # 75 realistic products, sorted by SKU
    products = [
        ('SKU-001', 'Wireless Mouse Pro',         'Peripherals',    29.99,  5),
        ('SKU-002', 'USB-C Hub 7-Port',            'Peripherals',    49.99,  3),
        ('SKU-003', 'Mechanical Keyboard RGB',     'Peripherals',    89.99,  2),
        ('SKU-004', 'HDMI Cable 2m',               'Cables',          9.99, 10),
        ('SKU-005', 'USB-A to USB-C Cable 1m',    'Cables',          7.99, 10),
        ('SKU-006', '24-inch Monitor FHD',         'Displays',      229.00,  1),
        ('SKU-007', '27-inch Monitor 4K',          'Displays',      449.00,  1),
        ('SKU-008', 'Laptop Stand Adjustable',     'Accessories',    34.99,  4),
        ('SKU-009', 'Desk Organizer Bamboo',       'Office',         22.50,  5),
        ('SKU-010', 'Wireless Headset Noise Cancel','Audio',         119.00,  2),
        ('SKU-011', 'Webcam 1080p HD',             'Peripherals',    59.99,  3),
        ('SKU-012', 'Docking Station USB-C',       'Peripherals',   149.99,  2),
        ('SKU-013', 'Ergonomic Office Chair',      'Furniture',     379.00,  1),
        ('SKU-014', 'Standing Desk Converter',     'Furniture',     249.00,  1),
        ('SKU-015', 'LED Desk Lamp',               'Lighting',       44.99,  3),
        ('SKU-016', 'Laptop Sleeve 15-inch',       'Accessories',    18.99,  5),
        ('SKU-017', 'Power Strip 6-Outlet',        'Power',          24.99,  4),
        ('SKU-018', 'Surge Protector 8-Outlet',    'Power',          39.99,  3),
        ('SKU-019', 'Cable Management Kit',        'Accessories',    15.99,  5),
        ('SKU-020', 'Monitor Arm Single',          'Accessories',    69.99,  2),
        ('SKU-021', 'Monitor Arm Dual',            'Accessories',    99.99,  2),
        ('SKU-022', 'USB 3.0 Flash Drive 64GB',    'Storage',        12.99, 10),
        ('SKU-023', 'Portable SSD 500GB',          'Storage',        79.99,  2),
        ('SKU-024', 'External HDD 1TB',            'Storage',        54.99,  2),
        ('SKU-025', 'Wireless Keyboard Slim',      'Peripherals',    49.99,  3),
        ('SKU-026', 'Bluetooth Numpad',            'Peripherals',    29.99,  5),
        ('SKU-027', 'Smart USB Power Strip',       'Power',          59.99,  2),
        ('SKU-028', 'Desk Mat Large',              'Accessories',    27.99,  4),
        ('SKU-029', 'Monitor Calibrator',          'Displays',      129.00,  2),
        ('SKU-030', 'Printer Ink Cartridge Black', 'Supplies',       14.99, 10),
        ('SKU-031', 'Printer Ink Cartridge Color', 'Supplies',       17.99, 10),
        ('SKU-032', 'Label Printer',               'Office',         89.99,  2),
        ('SKU-033', 'Paper Shredder',              'Office',         79.99,  1),
        ('SKU-034', 'Stapler Heavy Duty',          'Office',         22.99,  5),
        ('SKU-035', 'Tape Dispenser Desktop',      'Office',          8.99, 10),
        ('SKU-036', 'Whiteboard 36x24',            'Office',         59.99,  2),
        ('SKU-037', 'Dry Erase Markers Set',       'Supplies',        9.99, 10),
        ('SKU-038', 'File Folder Box',             'Office',         13.99,  5),
        ('SKU-039', 'Sticky Notes Multicolor',     'Supplies',        5.99, 20),
        ('SKU-040', 'Ballpoint Pens Pack 12',      'Supplies',        6.99, 10),
        ('SKU-041', 'Highlighter Set 6-Color',     'Supplies',        7.49, 10),
        ('SKU-042', 'Notebook Hardcover A4',       'Supplies',       12.99,  5),
        ('SKU-043', 'Binder 3-Ring 2-inch',        'Office',          5.99, 10),
        ('SKU-044', 'Calendar Planner 2025',       'Office',         18.99,  3),
        ('SKU-045', 'Ergonomic Wrist Rest',        'Accessories',    19.99,  5),
        ('SKU-046', 'Footrest Adjustable',         'Accessories',    44.99,  3),
        ('SKU-047', 'Lumbar Support Cushion',      'Accessories',    34.99,  4),
        ('SKU-048', 'Privacy Screen Filter 15"',   'Displays',       39.99,  3),
        ('SKU-049', 'Screen Cleaning Kit',         'Supplies',        9.99, 10),
        ('SKU-050', 'Cable Ties Pack 100',         'Accessories',     6.99, 10),
        ('SKU-051', 'Ethernet Cable Cat6 5m',      'Cables',         11.99,  5),
        ('SKU-052', 'Network Switch 8-Port',       'Networking',     39.99,  2),
        ('SKU-053', 'Wi-Fi Extender',              'Networking',     49.99,  2),
        ('SKU-054', 'Smartphone Stand Desktop',    'Accessories',    14.99,  5),
        ('SKU-055', 'Tablet Stand Adjustable',     'Accessories',    22.99,  4),
        ('SKU-056', 'Wireless Presenter Remote',   'Peripherals',    34.99,  3),
        ('SKU-057', 'Laser Pointer Green',         'Accessories',    19.99,  5),
        ('SKU-058', 'Video Conferencing Speakerphone','Audio',        89.99, 2),
        ('SKU-059', 'Noise Machine Desktop',       'Audio',          39.99,  3),
        ('SKU-060', 'USB Microphone',              'Audio',          79.99,  2),
        ('SKU-061', 'Ring Light 10-inch',          'Lighting',       49.99,  3),
        ('SKU-062', 'Background Green Screen',     'Accessories',    34.99,  2),
        ('SKU-063', 'Tripod Phone Mount',          'Accessories',    24.99,  4),
        ('SKU-064', 'Surge Protector Travel',      'Power',          22.99,  5),
        ('SKU-065', 'AA Batteries Pack 20',        'Power',          14.99, 10),
        ('SKU-066', 'AAA Batteries Pack 20',       'Power',          12.99, 10),
        ('SKU-067', 'Portable Battery Pack 20000mAh','Power',        59.99,  2),
        ('SKU-068', 'Wireless Charging Pad',       'Power',          29.99,  4),
        ('SKU-069', 'Multi-Port Charger 65W',      'Power',          49.99,  3),
        ('SKU-070', 'Ergonomic Vertical Mouse',    'Peripherals',    39.99,  3),
        ('SKU-071', 'Trackball Mouse',             'Peripherals',    54.99,  3),
        ('SKU-072', 'Drawing Tablet Small',        'Peripherals',    69.99,  2),
        ('SKU-073', 'Anti-Glare Screen Protector', 'Displays',       19.99,  5),
        ('SKU-074', 'Keyboard Cover Silicone',     'Accessories',     8.99, 10),
        ('SKU-075', 'VESA Mount Adapter Kit',      'Accessories',    15.99,  5),
    ]

    for r, prod in enumerate(products, 2):
        for c, val in enumerate(prod, 1):
            ws_pm.cell(row=r, column=c, value=val)

    # Column widths for ProductMaster
    ws_pm.column_dimensions['A'].width = 12
    ws_pm.column_dimensions['B'].width = 30
    ws_pm.column_dimensions['C'].width = 16
    ws_pm.column_dimensions['D'].width = 12
    ws_pm.column_dimensions['E'].width = 12

    # ── Orders sheet ──
    ws_ord = wb.create_sheet('Orders', 0)  # insert before ProductMaster

    ord_headers = ['SKU', 'Product Name', 'Category', 'List Price', 'Quantity', 'Discount %', 'Net Price']
    for col, h in enumerate(ord_headers, 1):
        cell = ws_ord.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # 100 order rows — column A has SKUs (from the 75 products pool), B-D empty, E qty, F discount, G empty
    import random
    random.seed(42)  # reproducibility

    # Realistic data: sales reps and their orders
    sku_list = [p[0] for p in products]

    # Generate 100 order rows
    for row in range(2, 102):
        sku = random.choice(sku_list)
        qty = random.randint(1, 50)
        discount = round(random.choice([0.0, 0.05, 0.10, 0.15, 0.20, 0.25]), 2)

        ws_ord.cell(row=row, column=1, value=sku)       # A: SKU
        # B (Product Name): intentionally empty
        # C (Category): intentionally empty
        # D (List Price): intentionally empty
        ws_ord.cell(row=row, column=5, value=qty)        # E: Quantity
        ws_ord.cell(row=row, column=6, value=discount)   # F: Discount %
        # G (Net Price): intentionally empty

    # Format column F as percentage
    for row in range(2, 102):
        ws_ord.cell(row=row, column=6).number_format = '0%'

    # Column widths for Orders
    ws_ord.column_dimensions['A'].width = 12
    ws_ord.column_dimensions['B'].width = 30
    ws_ord.column_dimensions['C'].width = 16
    ws_ord.column_dimensions['D'].width = 12
    ws_ord.column_dimensions['E'].width = 10
    ws_ord.column_dimensions['F'].width = 12
    ws_ord.column_dimensions['G'].width = 12

    # Freeze header row
    ws_ord.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Orders sheet: 100 rows, B/C/D/G columns empty (awaiting VLOOKUP formulas)')
    print(f'  ProductMaster sheet: 75 products sorted by SKU')


create_initial()
