"""
Initial Setup: Process 4 PDF invoices and update accounts_payable.ods
Task ID: osworld_multi_apps_doc_pdf_calc_005
Domain: libreoffice_calc (multi-app: PDF + Calc)

Creates:
  - /home/user/Desktop/invoices/ with 4 PDF invoices
  - /home/user/Desktop/accounts_payable.ods with starting balance row (no invoice rows)
"""

import os
import shlex
import subprocess
import time

WORKDIR = '/home/user'
DESKTOP = '/home/user/Desktop'
TASK_ID = 'osworld_multi_apps_doc_pdf_calc_005'
INVOICES_DIR = f'{DESKTOP}/invoices'
ODS_FILE = f'{DESKTOP}/accounts_payable.ods'
XLSX_TMP = f'/tmp/accounts_payable_tmp.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def run_cmd(cmd, timeout=60):
    """Run a shell command and return (returncode, stdout, stderr)."""
    result = subprocess.run(
        cmd, shell=True, capture_output=True, text=True, timeout=timeout
    )
    return result.returncode, result.stdout, result.stderr


def make_pdf_invoice(inv):
    """Create a single PDF invoice using fpdf2 with updated API."""
    from fpdf import FPDF
    from fpdf.enums import XPos, YPos

    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Vendor header
    pdf.set_font('Helvetica', 'B', 20)
    pdf.cell(0, 12, inv['vendor'], new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(0, 6, inv['vendor_addr'], new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(4)

    pdf.set_draw_color(0, 0, 0)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(4)

    pdf.set_font('Helvetica', 'B', 16)
    pdf.cell(0, 10, 'INVOICE', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(2)

    pdf.set_font('Helvetica', '', 11)
    pdf.cell(60, 7, 'Invoice Number:')
    pdf.set_font('Helvetica', 'B', 11)
    pdf.cell(0, 7, inv['invoice_no'], new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    pdf.set_font('Helvetica', '', 11)
    pdf.cell(60, 7, 'Invoice Date:')
    pdf.cell(0, 7, inv['date'], new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    pdf.cell(60, 7, 'Due Date:')
    pdf.cell(0, 7, inv['due_date'], new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    pdf.cell(60, 7, 'PO Number:')
    pdf.cell(0, 7, inv['po_number'], new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(6)

    pdf.set_font('Helvetica', 'B', 12)
    pdf.cell(0, 8, 'Bill To:', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font('Helvetica', '', 11)
    pdf.cell(0, 7, 'Acme Corporation', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.cell(0, 7, '1234 Business Park Drive', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.cell(0, 7, 'New York, NY 10001', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.cell(0, 7, 'Accounts Payable Department', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(6)

    # Table header
    pdf.set_fill_color(50, 50, 50)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 11)
    pdf.cell(90, 9, 'Description', border=1, fill=True)
    pdf.cell(25, 9, 'Qty', border=1, fill=True, align='C')
    pdf.cell(35, 9, 'Unit Price', border=1, fill=True, align='R')
    pdf.cell(35, 9, 'Total', border=1, fill=True, new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    # Table row
    pdf.set_fill_color(255, 255, 255)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Helvetica', '', 10)
    desc = inv['description'][:50]
    pdf.cell(90, 9, desc, border=1)
    pdf.cell(25, 9, str(inv['qty']), border=1, align='C')
    pdf.cell(35, 9, f"${inv['unit_price']:,.2f}", border=1, align='R')
    pdf.cell(35, 9, f"${inv['amount']:,.2f}", border=1, align='R',
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(4)

    # Totals
    pdf.set_font('Helvetica', '', 11)
    pdf.cell(150, 8, 'Subtotal:', align='R')
    pdf.cell(35, 8, f"${inv['amount']:,.2f}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='R')

    pdf.cell(150, 8, 'Tax (0%):', align='R')
    pdf.cell(35, 8, '$0.00', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='R')

    pdf.set_font('Helvetica', 'B', 12)
    pdf.cell(150, 9, 'TOTAL AMOUNT DUE:', align='R')
    pdf.cell(35, 9, f"${inv['amount']:,.2f}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='R')
    pdf.ln(8)

    pdf.set_font('Helvetica', 'I', 10)
    pdf.cell(0, 7, 'Payment Terms: Net 30 days from invoice date',
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.cell(0, 7, 'Please include invoice number on all remittances.',
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    return pdf


def create_invoices():
    """Create the invoices/ directory and 4 PDF invoice files."""
    os.makedirs(INVOICES_DIR, exist_ok=True)

    invoice_data = [
        {
            'filename': 'inv_001.pdf',
            'invoice_no': 'INV-2025-001',
            'vendor': 'Adobe Systems',
            'vendor_addr': '345 Park Avenue, San Jose, CA 95110',
            'date': '2025-01-10',
            'due_date': '2025-02-10',
            'amount': 599.88,
            'description': 'Adobe Creative Cloud - Enterprise License (Annual)',
            'qty': 1,
            'unit_price': 599.88,
            'po_number': 'PO-2025-0087',
        },
        {
            'filename': 'inv_002.pdf',
            'invoice_no': 'INV-2025-002',
            'vendor': 'AWS',
            'vendor_addr': '410 Terry Ave N, Seattle, WA 98109',
            'date': '2025-01-15',
            'due_date': '2025-02-15',
            'amount': 234.56,
            'description': 'Amazon Web Services - Cloud Infrastructure (January 2025)',
            'qty': 1,
            'unit_price': 234.56,
            'po_number': 'PO-2025-0091',
        },
        {
            'filename': 'inv_003.pdf',
            'invoice_no': 'INV-2025-003',
            'vendor': 'Slack Technologies',
            'vendor_addr': '500 Howard Street, San Francisco, CA 94105',
            'date': '2025-01-20',
            'due_date': '2025-02-20',
            'amount': 87.50,
            'description': 'Slack Pro Plan - Team Workspace Subscription (Monthly)',
            'qty': 1,
            'unit_price': 87.50,
            'po_number': 'PO-2025-0095',
        },
        {
            'filename': 'inv_004.pdf',
            'invoice_no': 'INV-2025-004',
            'vendor': 'Zoom Video',
            'vendor_addr': '55 Almaden Blvd, San Jose, CA 95113',
            'date': '2025-01-22',
            'due_date': '2025-02-22',
            'amount': 149.90,
            'description': 'Zoom Business Plan - Video Conferencing (Monthly)',
            'qty': 1,
            'unit_price': 149.90,
            'po_number': 'PO-2025-0098',
        },
    ]

    for inv in invoice_data:
        out_path = os.path.join(INVOICES_DIR, inv['filename'])
        pdf = make_pdf_invoice(inv)
        pdf.output(out_path)
        print(f'Created invoice: {out_path}')


def create_accounts_payable_ods():
    """Create accounts_payable.ods with starting balance row only (no invoice rows)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Accounts Payable'

    headers = ['Invoice_No', 'Vendor', 'Date', 'Amount', 'Status', 'Running_Total']
    header_fill = PatternFill(start_color='FF2E4057', end_color='FF2E4057', fill_type='solid')
    header_font = Font(name='Arial', bold=True, color='FFFFFFFF', size=11)
    thin_side = Side(style='thin', color='FF888888')
    header_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = header_border

    ws.row_dimensions[1].height = 22

    data_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )

    # Starting balance row
    starting_balance_row = [
        'BALANCE-FWD', 'Opening Balance', '2024-12-31', 0.00, 'Paid', 1250.00
    ]
    for col, val in enumerate(starting_balance_row, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.alignment = Alignment(horizontal='left' if col in [1, 2] else 'center')
        cell.border = data_border
        if col == 4 or col == 6:
            cell.number_format = '$#,##0.00'

    # 2 prior invoice rows (already reconciled)
    prior_invoices = [
        ['INV-2024-098', 'Microsoft', '2024-12-05', 450.00, 'Paid', 1700.00],
        ['INV-2024-112', 'Google Workspace', '2024-12-18', 216.50, 'Paid', 1916.50],
    ]
    for r, row_data in enumerate(prior_invoices, 3):
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = Alignment(horizontal='left' if col in [1, 2] else 'center')
            cell.border = data_border
            if col == 4 or col == 6:
                cell.number_format = '$#,##0.00'

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 16

    ws.freeze_panes = 'A2'

    # Save as xlsx, convert to ods
    wb.save(XLSX_TMP)
    print(f'Saved temporary xlsx: {XLSX_TMP}')

    convert_cmd = (
        f'libreoffice --headless --convert-to ods --outdir /tmp "{XLSX_TMP}"'
    )
    rc, out, err = run_cmd(convert_cmd, timeout=60)
    print(f'LibreOffice convert rc={rc}, out={out.strip()}, err={err.strip()}')

    converted = '/tmp/accounts_payable_tmp.ods'
    if os.path.exists(converted):
        import shutil
        shutil.copy(converted, ODS_FILE)
        print(f'Accounts payable ODS created: {ODS_FILE}')
    else:
        # Fallback
        import shutil
        shutil.copy(XLSX_TMP, ODS_FILE)
        print(f'Fallback: copied xlsx as ods: {ODS_FILE}')


def main():
    os.makedirs(DESKTOP, exist_ok=True)

    print('--- Creating invoice PDFs ---')
    create_invoices()

    print('--- Creating accounts_payable.ods ---')
    create_accounts_payable_ods()

    print('--- Verifying outputs ---')
    for fname in ['inv_001.pdf', 'inv_002.pdf', 'inv_003.pdf', 'inv_004.pdf']:
        path = os.path.join(INVOICES_DIR, fname)
        exists = os.path.exists(path)
        print(f'  {"OK" if exists else "MISSING"}: {path}')

    exists_ods = os.path.exists(ODS_FILE)
    print(f'  {"OK" if exists_ods else "MISSING"}: {ODS_FILE}')

    # GUI-ready startup
    time.sleep(1)
    launch_gui(f'nautilus "{DESKTOP}"', delay_sec=1.5)
    launch_gui(f'libreoffice --calc "{ODS_FILE}"', delay_sec=2.0)
    print('GUI_READY: launched Nautilus (Desktop) and LibreOffice Calc with accounts_payable.ods, DISPLAY=:0')


main()
