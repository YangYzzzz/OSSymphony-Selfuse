"""
FINAL REWARD SCRIPT - SUCCESS
Task: I need the Operating Margin filled in by calculating revenue minus cost of goods sold, operating expenses, and depreciation. Then in Sheet2 under column A called "Year_Margin", display the Year from Sheet 1 concatenated with "_" and the whole number portion of Operating Margin.
Generated: 2025-11-24 07:45:43
Status: success
Model: o3
Total Steps: 4
"""

import os
import math
import openpyxl

def verify_task(file_path):
    """Verify that Operating Margin is correctly calculated in Sheet1 and
    Year_Margin is correctly generated in Sheet2.

    Scoring (progressive – total 1.0):
        • Operating_Margin correctness 60%
        • Year_Margin correctness       40%
    """

    # ------------------------------------------------------------------
    # 1. Load workbook twice: once with formulas, once with data-only values
    # ------------------------------------------------------------------
    try:
        wb_formula = openpyxl.load_workbook(file_path, data_only=False)
        wb_value   = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"✗ Cannot open workbook: {e}")
        return 0.0

    # ------------------------------------------------------------------
    # 2. Basic sheet existence check (no points awarded, prerequisite)
    # ------------------------------------------------------------------
    if 'Sheet1' not in wb_formula.sheetnames or 'Sheet2' not in wb_formula.sheetnames:
        print("✗ Workbook must contain 'Sheet1' and 'Sheet2'")
        return 0.0

    sh1_f = wb_formula['Sheet1']  # with formulas
    sh1_v = wb_value['Sheet1']    # with values
    sh2_f = wb_formula['Sheet2']
    sh2_v = wb_value['Sheet2']

    # Determine number of data rows (exclude header)
    max_row = sh1_f.max_row
    data_rows = list(range(2, max_row + 1)) if max_row >= 2 else []
    print(f"Detected {len(data_rows)} data rows in Sheet1")
    if not data_rows:
        print("✗ No data rows found – nothing to verify")
        return 0.0

    # ------------------------------------------------------------------
    # 3. Scoring setup
    # ------------------------------------------------------------------
    operating_margin_score = 0.0  # up to 0.6
    year_margin_score      = 0.0  # up to 0.4
    per_margin_point       = 0.6 / len(data_rows)
    per_year_margin_point  = 0.4 / len(data_rows)

    # ------------------------------------------------------------------
    # 4. Row-by-row verification
    # ------------------------------------------------------------------
    for r in data_rows:
        year      = sh1_v.cell(row=r, column=1).value
        revenue   = sh1_v.cell(row=r, column=2).value
        cogs      = sh1_v.cell(row=r, column=3).value
        op_exp    = sh1_v.cell(row=r, column=4).value
        dep       = sh1_v.cell(row=r, column=5).value

        # Compute expected margin (ensure numeric conversion)
        try:
            expected_margin = (float(revenue) if revenue is not None else 0) \
                              - (float(cogs)  if cogs   is not None else 0) \
                              - (float(op_exp) if op_exp is not None else 0) \
                              - (float(dep)   if dep    is not None else 0)
        except Exception as e:
            print(f"  Row {r}: Non-numeric data – {e}")
            continue  # cannot verify this row

        # ---------------------------
        # 4a. Verify Operating_Margin
        # ---------------------------
        margin_val     = sh1_v.cell(row=r, column=6).value
        margin_formula = sh1_f.cell(row=r, column=6).value
        margin_ok = False

        # Case 1 – numeric value present & correct (±0.01 tolerance)
        if isinstance(margin_val, (int, float)) and abs(margin_val - expected_margin) < 0.01:
            margin_ok = True
            print(f"✓ Row {r}: Operating margin value correct ({margin_val})")

        # Case 2 – formula present using correct references
        if not margin_ok and isinstance(margin_formula, str) and margin_formula.startswith('='):
            refs = [f'B{r}', f'C{r}', f'D{r}', f'E{r}']
            formula_up = margin_formula.replace('$', '').upper()
            refs_ok = all(ref in formula_up for ref in refs)
            minus_ok = formula_up.count('-') >= 3
            if refs_ok and minus_ok:
                margin_ok = True
                print(f"✓ Row {r}: Operating margin formula correct ({margin_formula})")

        if margin_ok:
            operating_margin_score += per_margin_point
        else:
            print(f"✗ Row {r}: Operating margin incorrect")

        # ---------------------------
        # 4b. Verify Year_Margin in Sheet2
        # ---------------------------
        ym_val     = sh2_v.cell(row=r, column=1).value
        ym_formula = sh2_f.cell(row=r, column=1).value
        expected_ym = f"{year}_{int(math.floor(expected_margin))}"
        ym_ok = False

        # Direct value check
        if isinstance(ym_val, str) and ym_val == expected_ym:
            ym_ok = True
            print(f"✓ Row {r}: Year_Margin value correct ({ym_val})")

        # Formula pattern check
        if not ym_ok and isinstance(ym_formula, str) and ym_formula.startswith('='):
            up = ym_formula.replace('$', '').upper()
            year_refs   = [f'SHEET1!A{r}', f'A{r}']
            margin_refs = [f'SHEET1!F{r}', f'F{r}']
            if (any(ref in up for ref in year_refs) and
                any(ref in up for ref in margin_refs) and
                '"_"' in up):
                ym_ok = True
                print(f"✓ Row {r}: Year_Margin formula correct ({ym_formula})")

        if ym_ok:
            year_margin_score += per_year_margin_point
        else:
            print(f"✗ Row {r}: Year_Margin incorrect")

    # ------------------------------------------------------------------
    # 5. Final scoring
    # ------------------------------------------------------------------
    print(f"Operating margin score: {operating_margin_score:.3f} / 0.6")
    print(f"Year_Margin score: {year_margin_score:.3f} / 0.4")
    total = round(min(operating_margin_score + year_margin_score, 1.0), 3)
    print(f"Total score: {total}")
    return total


if __name__ == "__main__":
    # Adjust the path if the filename changes or resides elsewhere
    default_path = "/home/user/i_need_the_operating_margin_filled_in_by_calculating_revenue_minus_cost_of_goods_sold_operating_expe.xlsx"
    if not os.path.exists(default_path):
        print("✗ File not found – adjust the file_path in the script.")
        print("REWARD: 0.0")
    else:
        reward = verify_task(default_path)
        print(f"REWARD: {reward}")
