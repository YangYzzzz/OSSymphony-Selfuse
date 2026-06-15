"""
Initial Setup: KPI Dashboard with conditional formatting targets
Task ID: calc_gcv_036
Domain: libreoffice_calc

Creates a KPI Dashboard spreadsheet with metric labels and values in B3, B5, B7, B9.
No conditional formatting applied - that is the task for the agent.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_036'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KPI_Dashboard"

    # --- Title row ---
    ws.merge_cells("A1:B1")
    ws["A1"] = "Q1 2026 KPI Dashboard"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # --- Headers ---
    header_font = Font(name="Arial", size=11, bold=True)
    header_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
    for col, header in enumerate(["KPI Metric", "Value"], 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # --- KPI Data (rows 3, 5, 7, 9 with spacer rows) ---
    kpi_data = {
        3: ("Uptime %", 99.2),
        5: ("SLA Compliance", 87.5),
        7: ("Customer Satisfaction", 72.1),
        9: ("First Call Resolution", 94.8),
    }

    label_font = Font(name="Arial", size=11, bold=True)
    value_font = Font(name="Arial", size=12)

    for row, (label, value) in kpi_data.items():
        cell_a = ws.cell(row=row, column=1, value=label)
        cell_a.font = label_font
        cell_a.alignment = Alignment(vertical="center")

        cell_b = ws.cell(row=row, column=2, value=value)
        cell_b.font = value_font
        cell_b.number_format = '0.0'
        cell_b.alignment = Alignment(horizontal="center", vertical="center")

    # --- Spacer rows with light separator lines ---
    thin_border_bottom = Border(bottom=Side(style="thin", color="D9D9D9"))
    for spacer_row in [4, 6, 8]:
        for col in range(1, 3):
            ws.cell(row=spacer_row, column=col).border = thin_border_bottom

    # --- Additional context: Target and Status columns ---
    ws.cell(row=2, column=3, value="Target").font = header_font
    ws.cell(row=2, column=3).fill = header_fill
    ws.cell(row=2, column=3).alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=4, value="Period").font = header_font
    ws.cell(row=2, column=4).fill = header_fill
    ws.cell(row=2, column=4).alignment = Alignment(horizontal="center")

    targets = {3: (99.5, "Jan-Mar 2026"), 5: (95.0, "Jan-Mar 2026"),
               7: (85.0, "Jan-Mar 2026"), 9: (90.0, "Jan-Mar 2026")}
    for row, (target, period) in targets.items():
        ws.cell(row=row, column=3, value=target).number_format = '0.0'
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=period)
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")

    # --- Additional info row ---
    ws.cell(row=11, column=1, value="Last Updated:")
    ws.cell(row=11, column=1).font = Font(name="Arial", size=9, italic=True, color="808080")
    ws.cell(row=11, column=2, value="2026-03-31")
    ws.cell(row=11, column=2).font = Font(name="Arial", size=9, italic=True, color="808080")

    ws.cell(row=12, column=1, value="Report Owner:")
    ws.cell(row=12, column=1).font = Font(name="Arial", size=9, italic=True, color="808080")
    ws.cell(row=12, column=2, value="Elena Rodriguez, Operations Lead")
    ws.cell(row=12, column=2).font = Font(name="Arial", size=9, italic=True, color="808080")

    # --- Column widths ---
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16

    # --- Second sheet with historical data for realism ---
    ws2 = wb.create_sheet("Historical")
    ws2.cell(row=1, column=1, value="Month").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Uptime %").font = Font(bold=True)
    ws2.cell(row=1, column=3, value="SLA Compliance").font = Font(bold=True)
    ws2.cell(row=1, column=4, value="Cust. Satisfaction").font = Font(bold=True)
    ws2.cell(row=1, column=5, value="FCR").font = Font(bold=True)

    historical = [
        ("Oct 2025", 98.7, 91.2, 78.3, 89.1),
        ("Nov 2025", 99.1, 89.8, 76.5, 91.4),
        ("Dec 2025", 97.5, 85.3, 74.9, 92.7),
        ("Jan 2026", 99.4, 88.1, 73.2, 93.5),
        ("Feb 2026", 99.0, 86.9, 71.8, 94.2),
        ("Mar 2026", 99.2, 87.5, 72.1, 94.8),
    ]
    for r, row_data in enumerate(historical, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)
            if c > 1:
                ws2.cell(row=r, column=c).number_format = '0.0'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
