"""
Reward Script: Protect Sheet1 with password, allow selecting unlocked cells only, unlock B2:D20
Task ID: calc_ggf_039
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Sheet protection is enabled
  Component 2 (0.20): Password is set on the protection
  Component 3 (0.20): Select-locked-cells restricted, select-unlocked-cells allowed
  Component 4 (0.20): B2:D20 cells are unlocked (locked=False)
  Component 5 (0.15): Cells outside B2:D20 remain locked (locked=True)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_039'


def persist_app_state(domain: str):
    """Best-effort save for any open LibreOffice instance."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sheet1 must exist
    if 'Sheet1' not in wb.sheetnames:
        print("CRITICAL: Sheet1 not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Sheet1']

    # Component 1: Sheet protection is enabled (0.25 points)
    # This changes from sheet=False (initial) to sheet=True (golden)
    try:
        if ws.protection.sheet == True:
            print(f"PASS: Component 1 — Sheet protection is enabled (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — Sheet protection not enabled (sheet={ws.protection.sheet})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Password is set on the protection (0.20 points)
    # This changes from password=None (initial) to a non-empty password hash (golden)
    try:
        pwd = ws.protection.password
        if pwd is not None and pwd != '':
            print(f"PASS: Component 2 — Protection password is set (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 — No password set (password={repr(pwd)})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Selection permissions configured correctly (0.20 points)
    # selectLockedCells=True means users CANNOT select locked cells
    # selectUnlockedCells=False means users CAN select unlocked cells
    # In initial: both are False (no restrictions since sheet isn't protected)
    # In golden: selectLockedCells=True, selectUnlockedCells=False
    try:
        slc = ws.protection.selectLockedCells
        suc = ws.protection.selectUnlockedCells
        # selectLockedCells=True → locked cells cannot be selected
        # selectUnlockedCells=False → unlocked cells CAN be selected
        if slc == True and suc == False:
            print(f"PASS: Component 3 — Selection permissions correct: selectLockedCells={slc}, selectUnlockedCells={suc} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 3 — Expected selectLockedCells=True, selectUnlockedCells=False; got selectLockedCells={slc}, selectUnlockedCells={suc}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: B2:D20 cells are unlocked (0.20 points)
    # In initial: all cells have default locked=True
    # In golden: B2:D20 have locked=False
    try:
        unlocked_count = 0
        total_cells = 0
        for r in range(2, 21):  # rows 2-20
            for c in range(2, 5):  # columns B(2), C(3), D(4)
                total_cells += 1
                cell = ws.cell(row=r, column=c)
                if cell.protection.locked == False:
                    unlocked_count += 1

        if unlocked_count == total_cells:
            print(f"PASS: Component 4 — All {total_cells} cells in B2:D20 are unlocked (0.20 pts)")
            total_score += 0.20
        elif unlocked_count > 0:
            # Partial credit: proportional to how many cells are unlocked
            partial = 0.20 * (unlocked_count / total_cells)
            print(f"PARTIAL: Component 4 — {unlocked_count}/{total_cells} cells unlocked in B2:D20 ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — No cells in B2:D20 are unlocked ({unlocked_count}/{total_cells})")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Sheet protection active AND cells outside B2:D20 remain locked (0.15 points)
    # Compound check: anchored to the task change (protection enabled) so it fails on initial_env
    # In initial: locked=True but sheet not protected → compound check fails
    # In golden: locked=True AND sheet protected → compound check passes
    try:
        if ws.protection.sheet != True:
            print(f"FAIL: Component 5 — Sheet not protected, so locked cells have no effect")
        else:
            sample_coords = [
                'A1', 'B1', 'C1', 'D1', 'E1',  # header row
                'A2', 'E2', 'A10', 'E10',        # side columns in data range
                'A21', 'B21', 'C21', 'D21', 'E21',  # row below range
                'A25', 'B25',                      # further below
            ]
            locked_count = 0
            checked = 0
            for coord in sample_coords:
                cell = ws[coord]
                checked += 1
                if cell.protection.locked != False:
                    locked_count += 1

            if locked_count == checked:
                print(f"PASS: Component 5 — Sheet protected AND all {checked} sampled cells outside B2:D20 remain locked (0.15 pts)")
                total_score += 0.15
            else:
                unlocked_outside = checked - locked_count
                print(f"FAIL: Component 5 — {unlocked_outside}/{checked} sampled cells outside B2:D20 are unexpectedly unlocked")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
