"""
Initial Setup: Protect sheet with password and unlock input range B2:D20
Task ID: calc_ggf_039
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_039'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # --- Header row (row 1) - form labels ---
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    headers = {
        'A': 'Record #',
        'B': 'Product Name',
        'C': 'Unit Price ($)',
        'D': 'Quantity',
        'E': 'Line Total ($)',
    }
    for col_letter, header_text in headers.items():
        cell = ws[f'{col_letter}1']
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Column A: Row numbers (label column, not input) ---
    for r in range(2, 21):
        ws.cell(row=r, column=1, value=r - 1)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal='center')

    # --- B2:D20: Data entry area with some sample data ---
    sample_data = [
        ['Wireless Mouse', 29.99, 15],
        ['USB-C Hub', 45.50, 8],
        ['Mechanical Keyboard', 89.95, 12],
        ['Monitor Stand', 34.00, 20],
        ['Webcam HD Pro', 72.50, 5],
        ['Laptop Sleeve 15"', 19.99, 25],
        ['HDMI Cable 6ft', 12.49, 40],
        ['Desk Lamp LED', 38.75, 10],
        ['Noise-Cancel Headset', 129.00, 3],
        ['Surge Protector', 24.95, 18],
    ]
    for i, (product, price, qty) in enumerate(sample_data):
        row = i + 2
        ws.cell(row=row, column=2, value=product)
        ws.cell(row=row, column=3, value=price)
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        ws.cell(row=row, column=4, value=qty)

    # Rows 12-20 left empty for future data entry

    # --- Column E: Line Total formulas (computed, not input) ---
    for r in range(2, 21):
        ws.cell(row=r, column=5, value=f'=C{r}*D{r}')
        ws.cell(row=r, column=5).number_format = '#,##0.00'

    # --- Summary section (rows 21-24) - labels and formulas ---
    thin_border = Border(top=Side(style='thin', color='000000'))
    for col in range(1, 6):
        ws.cell(row=21, column=col).border = thin_border

    ws.cell(row=21, column=4, value='Subtotal:')
    ws.cell(row=21, column=4).font = Font(bold=True)
    ws.cell(row=21, column=4).alignment = Alignment(horizontal='right')
    ws.cell(row=21, column=5, value='=SUM(E2:E20)')
    ws.cell(row=21, column=5).number_format = '#,##0.00'
    ws.cell(row=21, column=5).font = Font(bold=True)

    ws.cell(row=22, column=4, value='Tax (8.5%):')
    ws.cell(row=22, column=4).font = Font(bold=True)
    ws.cell(row=22, column=4).alignment = Alignment(horizontal='right')
    ws.cell(row=22, column=5, value='=E21*0.085')
    ws.cell(row=22, column=5).number_format = '#,##0.00'

    ws.cell(row=23, column=4, value='Grand Total:')
    ws.cell(row=23, column=4).font = Font(bold=True, size=12)
    ws.cell(row=23, column=4).alignment = Alignment(horizontal='right')
    ws.cell(row=23, column=5, value='=E21+E22')
    ws.cell(row=23, column=5).number_format = '#,##0.00'
    ws.cell(row=23, column=5).font = Font(bold=True, size=12)

    ws.cell(row=25, column=1, value='Form Version: 2.1 | Last Updated: 2025-11-15')
    ws.cell(row=25, column=1).font = Font(italic=True, color='808080', size=9)

    # --- Column widths ---
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 16

    # --- NO protection applied (initial state is unprotected) ---

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
