"""
Initial Setup: Insert a new sheet using a template from a file
Task ID: calc_gsi_055
Domain: libreoffice_calc

Creates:
1. Main workbook calc_gsi_055.xlsx with two quarterly summary sheets
2. Template file monthly_report_template.xlsx with pre-formatted headers,
   formulas, and conditional formatting
3. Opens the main workbook in LibreOffice Calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_055'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'
TEMPLATE = f'{WORKDIR}/monthly_report_template.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_template():
    """Create the monthly report template file with formatting, formulas, and conditional formatting."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Report"

    # --- Styles ---
    header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2E5090", end_color="FF2E5090", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    subheader_fill = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")
    subheader_font = Font(name="Arial", size=10, bold=True)

    # --- Title Row ---
    ws.merge_cells("A1:G1")
    ws["A1"] = "Monthly Financial Report"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color="2E5090")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # --- Report Info ---
    ws["A2"] = "Report Period:"
    ws["A2"].font = Font(name="Arial", size=10, bold=True)
    ws["B2"] = "[Month Year]"
    ws["D2"] = "Prepared By:"
    ws["D2"].font = Font(name="Arial", size=10, bold=True)
    ws["E2"] = "[Name]"
    ws.row_dimensions[2].height = 20

    # --- Blank separator row ---
    ws.row_dimensions[3].height = 8

    # --- Headers (Row 4) ---
    headers = ["Category", "Budget", "Actual", "Variance", "Variance %", "Status", "Notes"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[4].height = 28

    # --- Category rows (5-19) with formulas ---
    categories = [
        "Revenue - Product Sales",
        "Revenue - Services",
        "Revenue - Subscriptions",
        "Cost of Goods Sold",
        "Gross Profit",
        "Operating Expenses - Salaries",
        "Operating Expenses - Marketing",
        "Operating Expenses - Rent",
        "Operating Expenses - Utilities",
        "Operating Expenses - Software",
        "Operating Expenses - Travel",
        "Total Operating Expenses",
        "Net Operating Income",
        "Tax Provision",
        "Net Income",
    ]

    for i, cat in enumerate(categories):
        row = 5 + i
        ws.cell(row=row, column=1, value=cat).font = Font(name="Arial", size=10)
        ws.cell(row=row, column=1).border = thin_border

        # Budget and Actual columns - placeholders
        for col in [2, 3]:
            cell = ws.cell(row=row, column=col)
            cell.number_format = '$#,##0.00'
            cell.border = thin_border

        # Variance formula: =C{row}-B{row}
        var_cell = ws.cell(row=row, column=4)
        var_cell.value = f'=C{row}-B{row}'
        var_cell.number_format = '$#,##0.00'
        var_cell.border = thin_border

        # Variance % formula: =IF(B{row}=0,"N/A",D{row}/B{row})
        pct_cell = ws.cell(row=row, column=5)
        pct_cell.value = f'=IF(B{row}=0,"N/A",D{row}/B{row})'
        pct_cell.number_format = '0.00%'
        pct_cell.border = thin_border

        # Status column - conditional formatting applied below
        ws.cell(row=row, column=6).border = thin_border
        # Notes column
        ws.cell(row=row, column=7).border = thin_border

    # Bold subtotal rows
    for subtotal_row_name in ["Gross Profit", "Total Operating Expenses", "Net Operating Income", "Net Income"]:
        for i, cat in enumerate(categories):
            if cat == subtotal_row_name:
                row = 5 + i
                for col in range(1, 8):
                    cell = ws.cell(row=row, column=col)
                    cell.font = Font(name="Arial", size=10, bold=True)
                    cell.fill = subheader_fill

    # --- Conditional Formatting: Highlight negative variances in red ---
    red_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
    red_font = Font(color="9C0006")
    ws.conditional_formatting.add(
        "D5:D19",
        CellIsRule(operator="lessThan", formula=["0"], fill=red_fill, font=red_font),
    )

    green_fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
    green_font = Font(color="006100")
    ws.conditional_formatting.add(
        "D5:D19",
        CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=green_fill, font=green_font),
    )

    # --- Footer row ---
    footer_row = 21
    ws.merge_cells(f"A{footer_row}:G{footer_row}")
    ws[f"A{footer_row}"] = "Confidential - For Internal Use Only"
    ws[f"A{footer_row}"].font = Font(name="Arial", size=8, italic=True, color="808080")
    ws[f"A{footer_row}"].alignment = Alignment(horizontal="center")

    # --- Column widths ---
    col_widths = {"A": 32, "B": 16, "C": 16, "D": 16, "E": 14, "F": 12, "G": 22}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A5"

    wb.save(TEMPLATE)
    print(f'Template file created: {TEMPLATE}')


def create_main_workbook():
    """Create the main workbook with two quarterly summary sheets."""
    wb = openpyxl.Workbook()

    # --- Sheet 1: Q1 Summary ---
    ws1 = wb.active
    ws1.title = "Q1 Summary"

    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF3A7D44", end_color="FF3A7D44", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    headers1 = ["Department", "Jan Revenue", "Feb Revenue", "Mar Revenue", "Q1 Total", "Target", "Achievement %"]
    for c, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    q1_data = [
        ["Engineering", 245000, 268000, 291000, None, 780000, None],
        ["Marketing", 182000, 195000, 203000, None, 560000, None],
        ["Sales", 410000, 438000, 465000, None, 1250000, None],
        ["Operations", 98000, 102000, 107000, None, 300000, None],
        ["Customer Success", 156000, 163000, 171000, None, 470000, None],
        ["Product", 128000, 135000, 142000, None, 390000, None],
        ["Finance", 67000, 71000, 74000, None, 200000, None],
        ["HR", 45000, 48000, 51000, None, 140000, None],
        ["Legal", 89000, 93000, 97000, None, 270000, None],
        ["R&D", 312000, 328000, 345000, None, 950000, None],
    ]

    for r, row_data in enumerate(q1_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c in [2, 3, 4, 5, 6]:
                cell.number_format = '$#,##0'
            if c == 7:
                cell.number_format = '0.0%'

        # Q1 Total formula
        ws1.cell(row=r, column=5).value = f'=SUM(B{r}:D{r})'
        # Achievement % formula
        ws1.cell(row=r, column=7).value = f'=IF(F{r}=0,"N/A",E{r}/F{r})'

    # Total row
    total_row = len(q1_data) + 2
    ws1.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", size=11, bold=True)
    ws1.cell(row=total_row, column=1).border = thin_border
    for c in range(2, 8):
        cell = ws1.cell(row=total_row, column=c)
        cell.border = thin_border
        cell.font = Font(name="Arial", size=11, bold=True)
        if c <= 6:
            col_letter = get_column_letter(c)
            cell.value = f'=SUM({col_letter}2:{col_letter}{total_row - 1})'
            cell.number_format = '$#,##0'
        if c == 7:
            cell.value = f'=IF(F{total_row}=0,"N/A",E{total_row}/F{total_row})'
            cell.number_format = '0.0%'

    col_widths1 = {"A": 20, "B": 15, "C": 15, "D": 15, "E": 15, "F": 15, "G": 16}
    for col_letter, width in col_widths1.items():
        ws1.column_dimensions[col_letter].width = width

    ws1.freeze_panes = "A2"

    # --- Sheet 2: Q2 Summary ---
    ws2 = wb.create_sheet("Q2 Summary")

    for c, h in enumerate(headers1, 1):
        h2 = h.replace("Jan", "Apr").replace("Feb", "May").replace("Mar", "Jun").replace("Q1", "Q2")
        cell = ws2.cell(row=1, column=c, value=h2)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    q2_data = [
        ["Engineering", 298000, 315000, 332000, None, 920000, None],
        ["Marketing", 210000, 225000, 238000, None, 650000, None],
        ["Sales", 478000, 502000, 531000, None, 1450000, None],
        ["Operations", 112000, 118000, 124000, None, 345000, None],
        ["Customer Success", 178000, 186000, 195000, None, 540000, None],
        ["Product", 148000, 156000, 164000, None, 455000, None],
        ["Finance", 76000, 80000, 84000, None, 235000, None],
        ["HR", 53000, 56000, 59000, None, 165000, None],
        ["Legal", 101000, 106000, 112000, None, 310000, None],
        ["R&D", 355000, 372000, 391000, None, 1080000, None],
    ]

    for r, row_data in enumerate(q2_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c in [2, 3, 4, 5, 6]:
                cell.number_format = '$#,##0'
            if c == 7:
                cell.number_format = '0.0%'

        ws2.cell(row=r, column=5).value = f'=SUM(B{r}:D{r})'
        ws2.cell(row=r, column=7).value = f'=IF(F{r}=0,"N/A",E{r}/F{r})'

    total_row2 = len(q2_data) + 2
    ws2.cell(row=total_row2, column=1, value="TOTAL").font = Font(name="Arial", size=11, bold=True)
    ws2.cell(row=total_row2, column=1).border = thin_border
    for c in range(2, 8):
        cell = ws2.cell(row=total_row2, column=c)
        cell.border = thin_border
        cell.font = Font(name="Arial", size=11, bold=True)
        if c <= 6:
            col_letter = get_column_letter(c)
            cell.value = f'=SUM({col_letter}2:{col_letter}{total_row2 - 1})'
            cell.number_format = '$#,##0'
        if c == 7:
            cell.value = f'=IF(F{total_row2}=0,"N/A",E{total_row2}/F{total_row2})'
            cell.number_format = '0.0%'

    for col_letter, width in col_widths1.items():
        ws2.column_dimensions[col_letter].width = width

    ws2.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Main workbook created: {OUTPUT}')


def main():
    create_template()
    create_main_workbook()

    # GUI-ready startup: open the main workbook in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


main()
