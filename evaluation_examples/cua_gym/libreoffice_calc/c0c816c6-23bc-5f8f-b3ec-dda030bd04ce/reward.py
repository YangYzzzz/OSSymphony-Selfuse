"""
Reward Script: Insert a new sheet from a template file (monthly_report_template.xlsx)
Task ID: calc_gsi_055
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): New 'Monthly Report' sheet exists in workbook
  Component 2 (0.25): Template header structure matches (row 4 headers with correct text)
  Component 3 (0.20): Template formulas present (variance/percentage formulas in D and E columns)
  Component 4 (0.15): Template formatting preserved (header fill color FF2E5090, bold, number formats)
  Component 5 (0.15): Template structural properties (merged cells, freeze panes, conditional formatting)
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_055'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: New 'Monthly Report' sheet exists (0.25 points)
    # Initial file has only 'Q1 Summary' and 'Q2 Summary'.
    # The task adds a 'Monthly Report' sheet from the template.
    try:
        sheet_names = wb.sheetnames
        # Check that a sheet with content matching the template exists.
        # The sheet name might vary slightly depending on how LO imports it,
        # so we look for a sheet that is NOT 'Q1 Summary' and NOT 'Q2 Summary'.
        new_sheets = [s for s in sheet_names if s not in ('Q1 Summary', 'Q2 Summary')]
        if len(new_sheets) >= 1:
            # Prefer exact name match
            if 'Monthly Report' in sheet_names:
                print(f"PASS: Component 1 — 'Monthly Report' sheet exists in workbook (0.25 pts)")
                total_score += 0.25
            else:
                # Accept any new sheet that was added from template
                print(f"PASS: Component 1 — New sheet '{new_sheets[0]}' found (likely from template) (0.25 pts)")
                total_score += 0.25
        else:
            print(f"FAIL: Component 1 — No new sheet found. Sheets: {sheet_names}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Find the template-derived sheet for further checks
    template_sheet = None
    try:
        if 'Monthly Report' in wb.sheetnames:
            template_sheet = wb['Monthly Report']
        else:
            # Try any new sheet
            new_sheets = [s for s in wb.sheetnames if s not in ('Q1 Summary', 'Q2 Summary')]
            if new_sheets:
                template_sheet = wb[new_sheets[0]]
    except Exception as e:
        print(f"ERROR: Could not locate template-derived sheet: {e}")

    if template_sheet is None:
        print("CRITICAL: No template-derived sheet found. Remaining components skipped.")
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    ws = template_sheet

    # Component 2: Template header structure in row 4 (0.25 points)
    # The template has 7 headers in row 4: Category, Budget, Actual, Variance,
    # Variance %, Status, Notes — with bold font and blue fill FF2E5090.
    try:
        expected_headers = ['Category', 'Budget', 'Actual', 'Variance', 'Variance %', 'Status', 'Notes']
        found_headers = []
        for col in range(1, 8):
            val = ws.cell(row=4, column=col).value
            if val is not None:
                found_headers.append(str(val).strip())
            else:
                found_headers.append(None)

        matches = sum(1 for i, eh in enumerate(expected_headers)
                      if i < len(found_headers) and found_headers[i] and
                      found_headers[i].lower() == eh.lower())

        if matches >= 6:
            print(f"PASS: Component 2 — Row 4 headers match template ({matches}/7) (0.25 pts)")
            total_score += 0.25
        elif matches >= 4:
            partial = round(0.25 * (matches / 7), 2)
            print(f"PARTIAL: Component 2 — Row 4 headers partially match ({matches}/7) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Row 4 headers don't match. Found: {found_headers}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Template formulas in D and E columns (0.20 points)
    # D5:D19 should have =C_-B_ formulas, E5:E19 should have =IF(B_=0,...) formulas
    try:
        formula_d_count = 0
        formula_e_count = 0
        for r in range(5, 20):
            d_val = ws.cell(row=r, column=4).value
            e_val = ws.cell(row=r, column=5).value
            if isinstance(d_val, str) and '=' in d_val:
                formula_d_count += 1
            if isinstance(e_val, str) and '=' in e_val:
                formula_e_count += 1

        total_formulas = formula_d_count + formula_e_count
        # Template has 15 D-formulas and 15 E-formulas = 30 total
        if total_formulas >= 24:
            print(f"PASS: Component 3 — Template formulas present (D:{formula_d_count}, E:{formula_e_count}) (0.20 pts)")
            total_score += 0.20
        elif total_formulas >= 12:
            partial = round(0.20 * (total_formulas / 30), 2)
            print(f"PARTIAL: Component 3 — Some template formulas present ({total_formulas}/30) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Too few formulas. D:{formula_d_count}, E:{formula_e_count}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Template formatting preserved (0.15 points)
    # Check: row 4 header cells have bold font AND blue fill (FF2E5090),
    # and B5:D19 have currency number format ($#,##0.00)
    try:
        fmt_score = 0.0

        # Sub-check 4a: Header bold + blue fill (0.075 pts)
        bold_blue_count = 0
        for col in range(1, 8):
            cell = ws.cell(row=4, column=col)
            is_bold = cell.font.bold
            try:
                fill_rgb = cell.fill.fgColor.rgb if cell.fill.patternType else None
            except:
                fill_rgb = None
            if is_bold and fill_rgb and '2E5090' in str(fill_rgb):
                bold_blue_count += 1

        if bold_blue_count >= 5:
            print(f"PASS: Component 4a — Header formatting (bold+blue): {bold_blue_count}/7")
            fmt_score += 0.075
        else:
            print(f"FAIL: Component 4a — Header formatting: only {bold_blue_count}/7 cells match")

        # Sub-check 4b: Currency number formats in B5:D5 (0.075 pts)
        currency_count = 0
        for r in range(5, 10):
            for c in range(2, 5):
                nf = ws.cell(row=r, column=c).number_format
                if nf and '$' in str(nf):
                    currency_count += 1

        if currency_count >= 10:
            print(f"PASS: Component 4b — Currency number formats found: {currency_count}")
            fmt_score += 0.075
        else:
            print(f"FAIL: Component 4b — Currency format cells: {currency_count} (expected >= 10)")

        if fmt_score > 0:
            total_score += fmt_score
            print(f"  Component 4 total: {fmt_score} pts")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Template structural properties (0.15 points)
    # Check: merged cells (A1:G1, A21:G21), freeze panes (A5),
    # and conditional formatting on D5:D19
    try:
        struct_score = 0.0

        # Sub-check 5a: Merged cells present (0.05 pts)
        merged = list(ws.merged_cells.ranges)
        merged_strs = [str(m) for m in merged]
        has_title_merge = any('A1' in m and 'G1' in m for m in merged_strs)
        has_footer_merge = any('A21' in m and 'G21' in m for m in merged_strs)
        if has_title_merge or has_footer_merge:
            print(f"PASS: Component 5a — Merged cells found: {merged_strs}")
            struct_score += 0.05
        else:
            print(f"FAIL: Component 5a — Expected merged cells A1:G1 and/or A21:G21, found: {merged_strs}")

        # Sub-check 5b: Freeze panes at A5 (0.05 pts)
        if ws.freeze_panes and str(ws.freeze_panes) == 'A5':
            print(f"PASS: Component 5b — Freeze panes at A5")
            struct_score += 0.05
        else:
            print(f"FAIL: Component 5b — Freeze panes: {ws.freeze_panes} (expected A5)")

        # Sub-check 5c: Conditional formatting exists (0.05 pts)
        cf_list = list(ws.conditional_formatting)
        if len(cf_list) >= 1:
            print(f"PASS: Component 5c — Conditional formatting rules found: {len(cf_list)}")
            struct_score += 0.05
        else:
            print(f"FAIL: Component 5c — No conditional formatting found")

        if struct_score > 0:
            total_score += struct_score
            print(f"  Component 5 total: {struct_score} pts")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
