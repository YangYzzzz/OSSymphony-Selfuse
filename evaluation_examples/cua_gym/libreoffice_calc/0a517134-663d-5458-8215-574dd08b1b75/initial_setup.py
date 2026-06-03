"""
Initial Setup: Weekly work schedule for a retail store
Task ID: calc_grs_021
Domain: libreoffice_calc

Creates a spreadsheet with 7 day-sheets, each containing a grid of 12 employees
and 16 hourly time slots (7am-10pm). Cells contain raw shift values (OFF, OPEN,
MID, CLOSE, TRAINING) but NO data validation, NO conditional formatting,
NO count formulas, NO total hours column, and NO freeze panes.
"""

import os
import shlex
import subprocess
import time
import random
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_021'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

EMPLOYEES = [
    'Sarah Chen', 'Marcus Johnson', 'Priya Patel', 'James O\'Brien',
    'Lucia Fernandez', 'David Kim', 'Aisha Williams', 'Ryan Thompson',
    'Elena Kowalski', 'Carlos Rivera', 'Megan Foster', 'Jamal Washington'
]

DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

TIME_SLOTS = [
    '7am', '8am', '9am', '10am', '11am', '12pm',
    '1pm', '2pm', '3pm', '4pm', '5pm', '6pm',
    '7pm', '8pm', '9pm', '10pm'
]

SHIFT_TYPES = ['OFF', 'OPEN', 'MID', 'CLOSE', 'TRAINING']

# Realistic shift patterns per employee role
SHIFT_PATTERNS = {
    # Full-time openers (early shifts)
    0: ['OPEN', 'OPEN', 'OPEN', 'OPEN', 'OPEN', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF'],
    1: ['OPEN', 'OPEN', 'OPEN', 'OPEN', 'OPEN', 'OPEN', 'OPEN', 'OPEN', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF'],
    # Full-time mid shifts
    2: ['OFF', 'OFF', 'MID', 'MID', 'MID', 'MID', 'MID', 'MID', 'MID', 'MID', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF'],
    3: ['OFF', 'OFF', 'OFF', 'MID', 'MID', 'MID', 'MID', 'MID', 'MID', 'MID', 'MID', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF'],
    # Full-time closers (late shifts)
    4: ['OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'CLOSE', 'CLOSE', 'CLOSE', 'CLOSE', 'CLOSE', 'CLOSE', 'CLOSE', 'CLOSE'],
    5: ['OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'CLOSE', 'CLOSE', 'CLOSE', 'CLOSE', 'CLOSE', 'CLOSE', 'CLOSE', 'CLOSE', 'OFF'],
    # Part-time various
    6: ['OFF', 'OPEN', 'OPEN', 'OPEN', 'OPEN', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF'],
    7: ['OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'MID', 'MID', 'MID', 'MID', 'MID', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF'],
    8: ['OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'CLOSE', 'CLOSE', 'CLOSE', 'CLOSE', 'OFF', 'OFF', 'OFF'],
    # Training + mixed
    9: ['TRAINING', 'TRAINING', 'TRAINING', 'TRAINING', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF'],
    10: ['OFF', 'OFF', 'OFF', 'OPEN', 'OPEN', 'OPEN', 'OPEN', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF'],
    11: ['OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'MID', 'MID', 'MID', 'MID', 'MID', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF', 'OFF'],
}


def get_shift(emp_idx, day_idx, slot_idx):
    """Generate a realistic shift value based on employee and day."""
    random.seed(emp_idx * 100 + day_idx * 16 + slot_idx)

    # Weekend (Sat/Sun) - some employees off
    if day_idx >= 5 and emp_idx in [0, 3, 9]:
        return 'OFF'

    base = SHIFT_PATTERNS[emp_idx][slot_idx]

    # Add some variation per day
    if random.random() < 0.1 and base != 'OFF':
        return random.choice(['OFF', base, base])
    if random.random() < 0.05 and base == 'OFF':
        return 'TRAINING'

    return base


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    header_font = Font(name='Arial', size=11, bold=True)
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_font_white = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for day_idx, day_name in enumerate(DAYS):
        if day_idx == 0:
            ws = wb.active
            ws.title = day_name
        else:
            ws = wb.create_sheet(day_name)

        # Column A header: Employee
        ws.cell(row=1, column=1, value='Employee')
        ws['A1'].font = header_font_white
        ws['A1'].fill = header_fill
        ws['A1'].alignment = center_align
        ws['A1'].border = thin_border
        ws.column_dimensions['A'].width = 20

        # Time slot headers (columns B through Q = 16 columns)
        for col_idx, slot in enumerate(TIME_SLOTS, 2):
            cell = ws.cell(row=1, column=col_idx, value=slot)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = 10

        # Employee rows (rows 2-13)
        for emp_idx, emp_name in enumerate(EMPLOYEES):
            row = emp_idx + 2
            name_cell = ws.cell(row=row, column=1, value=emp_name)
            name_cell.font = Font(name='Arial', size=10, bold=True)
            name_cell.alignment = Alignment(vertical='center')
            name_cell.border = thin_border

            for slot_idx in range(len(TIME_SLOTS)):
                col = slot_idx + 2
                shift = get_shift(emp_idx, day_idx, slot_idx)
                cell = ws.cell(row=row, column=col, value=shift)
                cell.alignment = center_align
                cell.border = thin_border
                cell.font = Font(name='Arial', size=9)

        # Row height for data rows
        for r in range(1, len(EMPLOYEES) + 2):
            ws.row_dimensions[r].height = 22

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
