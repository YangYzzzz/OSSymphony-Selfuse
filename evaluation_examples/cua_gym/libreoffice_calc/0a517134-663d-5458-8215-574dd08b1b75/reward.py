"""
Reward Script: Weekly Work Schedule for Retail Store
Task ID: calc_grs_021
Domain: libreoffice_calc

Scoring Rubric (5 components, total 1.0):
  1. Staff Count row with COUNTIF formulas (0.25)
  2. Total Hours column with COUNTIF formulas (0.25)
  3. Freeze panes at B2 on all sheets (0.15)
  4. Data validation (dropdown list) on shift cells (0.15)
  5. Conditional formatting with 5 color rules (0.20)

All components verify task-introduced changes only (absent in initial_env).
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_021'

DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
VALID_SHIFTS = {'OFF', 'OPEN', 'MID', 'CLOSE', 'TRAINING'}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Must have 7 day sheets
    for day in DAYS:
        if day not in wb.sheetnames:
            print(f"CRITICAL: Missing sheet '{day}'")
            print("REWARD: 0.0")
            return 0.0

    # =========================================================================
    # Component 1: Staff Count row with COUNTIF formulas (0.25 points)
    # The golden file has row 14 with A14="Staff Count" and B14:Q14 containing
    # COUNTIF formulas counting non-OFF employees per hour slot.
    # This row does NOT exist in the initial file.
    # =========================================================================
    try:
        sheets_with_count_row = 0
        for day in DAYS:
            ws = wb[day]
            # Check that there is a "Staff Count" label in column A, row 14+
            found_label = False
            count_row = None
            for r in range(14, ws.max_row + 1):
                val = ws.cell(r, 1).value
                if val and 'staff' in str(val).lower() and 'count' in str(val).lower():
                    found_label = True
                    count_row = r
                    break
            if not found_label:
                # Also check if it's labeled differently but has COUNTIF in row 14
                if ws.cell(14, 1).value is not None:
                    count_row = 14
                    found_label = True

            if found_label and count_row is not None:
                # Check that at least some columns have COUNTIF formulas
                countif_count = 0
                for c in range(2, 18):  # B through Q (cols 2-17)
                    cell_val = ws.cell(count_row, c).value
                    if cell_val and isinstance(cell_val, str) and 'COUNTIF' in cell_val.upper():
                        countif_count += 1
                if countif_count >= 10:  # at least 10 of 16 time slots
                    sheets_with_count_row += 1

        if sheets_with_count_row == 7:
            print(f"PASS: Component 1 -- Staff Count row with COUNTIF formulas on all 7 sheets (0.25 pts)")
            total_score += 0.25
        elif sheets_with_count_row > 0:
            partial = round(0.25 * (sheets_with_count_row / 7), 2)
            print(f"PARTIAL: Component 1 -- Staff Count row found on {sheets_with_count_row}/7 sheets ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 -- No Staff Count row with COUNTIF formulas found on any sheet")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # =========================================================================
    # Component 2: Total Hours column with COUNTIF formulas (0.25 points)
    # The golden file has column R (18) with R1="Total Hours" and R2:R13
    # containing COUNTIF formulas counting non-OFF hours per employee.
    # This column does NOT exist in the initial file (initial max_col=17).
    # =========================================================================
    try:
        sheets_with_hours_col = 0
        for day in DAYS:
            ws = wb[day]
            # Find the "Total Hours" header in row 1
            hours_col = None
            for c in range(1, ws.max_column + 1):
                val = ws.cell(1, c).value
                if val and 'total' in str(val).lower() and 'hour' in str(val).lower():
                    hours_col = c
                    break

            if hours_col is not None:
                # Check that employee rows (2-13) have COUNTIF formulas
                countif_count = 0
                for r in range(2, 14):  # rows 2-13 (12 employees)
                    cell_val = ws.cell(r, hours_col).value
                    if cell_val and isinstance(cell_val, str) and 'COUNTIF' in cell_val.upper():
                        countif_count += 1
                if countif_count >= 8:  # at least 8 of 12 employees
                    sheets_with_hours_col += 1

        if sheets_with_hours_col == 7:
            print(f"PASS: Component 2 -- Total Hours column with COUNTIF formulas on all 7 sheets (0.25 pts)")
            total_score += 0.25
        elif sheets_with_hours_col > 0:
            partial = round(0.25 * (sheets_with_hours_col / 7), 2)
            print(f"PARTIAL: Component 2 -- Total Hours column found on {sheets_with_hours_col}/7 sheets ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- No Total Hours column with COUNTIF formulas found on any sheet")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # =========================================================================
    # Component 3: Freeze panes at B2 on all sheets (0.15 points)
    # The golden file has freeze_panes="B2" on all 7 day sheets.
    # The initial file has freeze_panes=None on all sheets.
    # =========================================================================
    try:
        sheets_with_freeze = 0
        for day in DAYS:
            ws = wb[day]
            if ws.freeze_panes is not None and str(ws.freeze_panes) == 'B2':
                sheets_with_freeze += 1

        if sheets_with_freeze == 7:
            print(f"PASS: Component 3 -- Freeze panes at B2 on all 7 sheets (0.15 pts)")
            total_score += 0.15
        elif sheets_with_freeze > 0:
            # Accept any freeze pane that freezes at least the first column
            partial = round(0.15 * (sheets_with_freeze / 7), 2)
            print(f"PARTIAL: Component 3 -- Freeze panes at B2 on {sheets_with_freeze}/7 sheets ({partial} pts)")
            total_score += partial
        else:
            # Also accept other reasonable freeze locations (e.g. B1 freezes col A)
            sheets_with_any_freeze = 0
            for day in DAYS:
                ws = wb[day]
                if ws.freeze_panes is not None:
                    sheets_with_any_freeze += 1
            if sheets_with_any_freeze >= 5:
                print(f"PARTIAL: Component 3 -- Non-standard freeze panes on {sheets_with_any_freeze}/7 sheets (0.08 pts)")
                total_score += 0.08
            else:
                print(f"FAIL: Component 3 -- No freeze panes found (initial has none)")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # =========================================================================
    # Component 4: Data validation dropdown list on shift cells (0.15 points)
    # The golden file has a list validation with "OFF,OPEN,MID,CLOSE,TRAINING"
    # applied to B2:Q13 on all sheets. Initial has NO data validations.
    # =========================================================================
    try:
        sheets_with_validation = 0
        for day in DAYS:
            ws = wb[day]
            if ws.data_validations and len(ws.data_validations.dataValidation) > 0:
                for dv in ws.data_validations.dataValidation:
                    if dv.type == 'list' and dv.formula1:
                        # Check that the formula contains the expected shift types
                        formula_upper = str(dv.formula1).upper()
                        has_shifts = all(s in formula_upper for s in ['OFF', 'OPEN', 'MID', 'CLOSE', 'TRAINING'])
                        if has_shifts:
                            sheets_with_validation += 1
                            break

        if sheets_with_validation == 7:
            print(f"PASS: Component 4 -- Data validation dropdown on all 7 sheets (0.15 pts)")
            total_score += 0.15
        elif sheets_with_validation > 0:
            partial = round(0.15 * (sheets_with_validation / 7), 2)
            print(f"PARTIAL: Component 4 -- Data validation found on {sheets_with_validation}/7 sheets ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 -- No data validation with shift types found on any sheet")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # =========================================================================
    # Component 5: Conditional formatting with 5 color-coded shift rules (0.20 points)
    # The golden file has conditional formatting on B2:Q13 with rules for
    # OPEN (green), MID (yellow), CLOSE (orange), TRAINING (blue), OFF (gray).
    # The initial file has NO conditional formatting.
    # =========================================================================
    try:
        sheets_with_cf = 0
        for day in DAYS:
            ws = wb[day]
            cf_rules_list = list(ws.conditional_formatting)
            if len(cf_rules_list) > 0:
                # Count how many shift-type rules exist
                shift_rules_found = 0
                for cf in cf_rules_list:
                    for rule in cf.rules:
                        if rule.formula:
                            for f in rule.formula:
                                f_upper = str(f).upper().replace('"', '').replace("'", '')
                                if f_upper in VALID_SHIFTS:
                                    shift_rules_found += 1
                if shift_rules_found >= 3:  # at least 3 of 5 shift types
                    sheets_with_cf += 1

        if sheets_with_cf == 7:
            print(f"PASS: Component 5 -- Conditional formatting rules on all 7 sheets (0.20 pts)")
            total_score += 0.20
        elif sheets_with_cf > 0:
            partial = round(0.20 * (sheets_with_cf / 7), 2)
            print(f"PARTIAL: Component 5 -- Conditional formatting found on {sheets_with_cf}/7 sheets ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 -- No conditional formatting with shift-type rules found on any sheet")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
