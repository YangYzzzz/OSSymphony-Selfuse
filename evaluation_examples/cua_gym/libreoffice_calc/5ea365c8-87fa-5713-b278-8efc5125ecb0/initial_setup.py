"""
Initial Setup: Sales Territory Rep Assignment
Task ID: calc_sales_territory_rep_assign_010
Domain: libreoffice_calc

Creates a workbook with:
  - RepList sheet: 25 reps with Rep ID, Name, empty Territory (C), Individual Quota, empty Coverage % (E)
  - RepDirectory sheet: Rep ID to Territory mapping
  - TerritoryTargets sheet: 5 territories with quota targets
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_territory_rep_assign_010'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # -------------------------------------------------------
    # Sheet 1: RepList
    # -------------------------------------------------------
    ws_rep = wb.active
    ws_rep.title = 'RepList'

    # Headers
    headers = ['Rep ID', 'Rep Name', 'Territory', 'Individual Quota', 'Territory Coverage %']
    for col, h in enumerate(headers, 1):
        cell = ws_rep.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')

    # 25 reps spread across 5 territories (C and E columns are EMPTY)
    reps_data = [
        # (Rep ID, Rep Name, Individual Quota)  — Territory and Coverage % left blank
        ('R001', 'Sarah Chen',       520000),
        ('R002', 'Marcus Johnson',   480000),
        ('R003', 'Emily Rodriguez',  610000),
        ('R004', 'David Kim',        445000),
        ('R005', 'Jennifer Walsh',   530000),
        ('R006', 'Robert Patel',     390000),
        ('R007', 'Amanda Torres',    415000),
        ('R008', 'Michael Nguyen',   460000),
        ('R009', 'Lisa Hernandez',   375000),
        ('R010', 'James O\'Brien',   490000),
        ('R011', 'Stephanie Clark',  720000),
        ('R012', 'Brian Mitchell',   650000),
        ('R013', 'Rachel Anderson',  580000),
        ('R014', 'Kevin Thompson',   790000),
        ('R015', 'Nicole Davis',     620000),
        ('R016', 'Christopher Lee',  310000),
        ('R017', 'Megan Wilson',     280000),
        ('R018', 'Andrew Martin',    295000),
        ('R019', 'Brittany Jackson', 320000),
        ('R020', 'Daniel White',     265000),
        ('R021', 'Patricia Brown',   680000),
        ('R022', 'Steven Taylor',    710000),
        ('R023', 'Lauren Harris',    640000),
        ('R024', 'Mark Robinson',    595000),
        ('R025', 'Sandra Lewis',     730000),
    ]

    for row_idx, (rep_id, rep_name, quota) in enumerate(reps_data, 2):
        ws_rep.cell(row=row_idx, column=1, value=rep_id)
        ws_rep.cell(row=row_idx, column=2, value=rep_name)
        # Column C (Territory) — intentionally left EMPTY
        # Column D (Individual Quota)
        cell_d = ws_rep.cell(row=row_idx, column=4, value=quota)
        cell_d.number_format = '$#,##0'
        # Column E (Territory Coverage %) — intentionally left EMPTY

    # Column widths for readability
    ws_rep.column_dimensions['A'].width = 10
    ws_rep.column_dimensions['B'].width = 22
    ws_rep.column_dimensions['C'].width = 18
    ws_rep.column_dimensions['D'].width = 18
    ws_rep.column_dimensions['E'].width = 22
    ws_rep.freeze_panes = 'A2'

    # -------------------------------------------------------
    # Sheet 2: RepDirectory
    # Rep ID to Territory mapping
    # -------------------------------------------------------
    ws_dir = wb.create_sheet('RepDirectory')

    dir_headers = ['Rep ID', 'Territory']
    for col, h in enumerate(dir_headers, 1):
        cell = ws_dir.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FF70AD47', end_color='FF70AD47', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')

    # Mapping: 5 reps per territory, 5 territories
    rep_territory_map = [
        ('R001', 'North East'),
        ('R002', 'North East'),
        ('R003', 'North East'),
        ('R004', 'North East'),
        ('R005', 'North East'),
        ('R006', 'South East'),
        ('R007', 'South East'),
        ('R008', 'South East'),
        ('R009', 'South East'),
        ('R010', 'South East'),
        ('R011', 'Mid West'),
        ('R012', 'Mid West'),
        ('R013', 'Mid West'),
        ('R014', 'Mid West'),
        ('R015', 'Mid West'),
        ('R016', 'South West'),
        ('R017', 'South West'),
        ('R018', 'South West'),
        ('R019', 'South West'),
        ('R020', 'South West'),
        ('R021', 'West Coast'),
        ('R022', 'West Coast'),
        ('R023', 'West Coast'),
        ('R024', 'West Coast'),
        ('R025', 'West Coast'),
    ]

    for row_idx, (rep_id, territory) in enumerate(rep_territory_map, 2):
        ws_dir.cell(row=row_idx, column=1, value=rep_id)
        ws_dir.cell(row=row_idx, column=2, value=territory)

    ws_dir.column_dimensions['A'].width = 10
    ws_dir.column_dimensions['B'].width = 18

    # -------------------------------------------------------
    # Sheet 3: TerritoryTargets
    # 5 territories with quota targets
    # -------------------------------------------------------
    ws_tgt = wb.create_sheet('TerritoryTargets')

    tgt_headers = ['Territory', 'Target Quota']
    for col, h in enumerate(tgt_headers, 1):
        cell = ws_tgt.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFED7D31', end_color='FFED7D31', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')

    # Territory targets:
    # North East: $2,500,000 — 5 reps total: 520+480+610+445+530 = 2,585,000 (>= target: green)
    # South East: $1,800,000 — 5 reps total: 390+415+460+375+490 = 2,130,000 (>= target: green)
    # Mid West:   $3,500,000 — 5 reps total: 720+650+580+790+620 = 3,360,000 (< target: red)
    # South West: $1,200,000 — 5 reps total: 310+280+295+320+265 = 1,470,000 (>= target: green)
    # West Coast: $2,800,000 — 5 reps total: 680+710+640+595+730 = 3,355,000 (>= target: green)
    territory_targets = [
        ('North East', 2500000),
        ('South East', 1800000),
        ('Mid West',   3500000),
        ('South West', 1200000),
        ('West Coast', 2800000),
    ]

    for row_idx, (territory, target) in enumerate(territory_targets, 2):
        ws_tgt.cell(row=row_idx, column=1, value=territory)
        cell_b = ws_tgt.cell(row=row_idx, column=2, value=target)
        cell_b.number_format = '$#,##0'

    ws_tgt.column_dimensions['A'].width = 18
    ws_tgt.column_dimensions['B'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  RepList: 25 reps with empty Territory (C) and Coverage% (E) columns')
    print(f'  RepDirectory: 25 rep-to-territory mappings')
    print(f'  TerritoryTargets: 5 territories with quota targets')


create_initial()
