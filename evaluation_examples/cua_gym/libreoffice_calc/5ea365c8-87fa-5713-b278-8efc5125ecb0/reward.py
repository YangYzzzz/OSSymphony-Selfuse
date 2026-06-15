"""
Reward Script: Assign reps to territories with VLOOKUP and calculate coverage %
Task ID: calc_sales_territory_rep_assign_010
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.40): VLOOKUP formulas in RepList column C (rows 2-26) pulling territory from RepDirectory
  - Component 2 (0.35): Coverage % formulas in RepList column E (rows 2-26) dividing individual quota by territory target via VLOOKUP
  - Component 3 (0.25): Conditional formatting on E2:E26 with red fill for < 1.0 and green fill for >= 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_sales_territory_rep_assign_010'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: RepList sheet must exist
    if 'RepList' not in wb.sheetnames:
        print("CRITICAL: 'RepList' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['RepList']

    # Component 1: VLOOKUP formulas in column C (rows 2-26) for territory assignment (0.40 points)
    # The formula should be of the form =VLOOKUP(A<n>, RepDirectory!..., 2, FALSE)
    # This FAILS on initial (all None) and PASSES on golden (all have VLOOKUP formulas)
    try:
        vlookup_count = 0
        vlookup_total = 25  # rows 2 through 26

        for row_num in range(2, 27):
            cell_val = ws.cell(row=row_num, column=3).value  # Column C
            if cell_val is not None and isinstance(cell_val, str):
                # Check that it's a VLOOKUP referencing RepDirectory column B
                val_upper = cell_val.upper().replace(' ', '')
                if ('VLOOKUP' in val_upper and
                    'REPDIRECTORY' in val_upper and
                    ',2,' in val_upper):
                    vlookup_count += 1

        if vlookup_count == vlookup_total:
            print(f"PASS: Component 1 — All {vlookup_total} VLOOKUP formulas present in column C ({0.40} pts)")
            total_score += 0.40
        elif vlookup_count > 0:
            partial = round(0.40 * vlookup_count / vlookup_total, 4)
            print(f"PARTIAL: Component 1 — {vlookup_count}/{vlookup_total} VLOOKUP formulas in column C ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No VLOOKUP formulas found in column C (expected {vlookup_total})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Territory Coverage % formulas in column E (rows 2-26) (0.35 points)
    # The formula divides individual quota (D column) by territory target via VLOOKUP from TerritoryTargets
    # Pattern: =D<n>/VLOOKUP(C<n>, TerritoryTargets!..., 2, FALSE)
    # This FAILS on initial (all None) and PASSES on golden
    try:
        coverage_count = 0
        coverage_total = 25  # rows 2 through 26

        for row_num in range(2, 27):
            cell_val = ws.cell(row=row_num, column=5).value  # Column E
            if cell_val is not None and isinstance(cell_val, str):
                val_upper = cell_val.upper().replace(' ', '')
                # Must reference D column (the individual quota) and VLOOKUP from TerritoryTargets
                if ('VLOOKUP' in val_upper and
                    'TERRITORYTARGETS' in val_upper and
                    ',2,' in val_upper):
                    coverage_count += 1

        if coverage_count == coverage_total:
            print(f"PASS: Component 2 — All {coverage_total} coverage % formulas present in column E ({0.35} pts)")
            total_score += 0.35
        elif coverage_count > 0:
            partial = round(0.35 * coverage_count / coverage_total, 4)
            print(f"PARTIAL: Component 2 — {coverage_count}/{coverage_total} coverage formulas in column E ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No coverage % formulas found in column E (expected {coverage_total})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Conditional formatting on E2:E26 with red for < 1.0 and green for >= 1.0 (0.25 points)
    # This FAILS on initial (no CF rules) and PASSES on golden
    try:
        cf_rules = ws.conditional_formatting
        red_rule_count = 0    # counts valid red-fill lessThan rules
        green_rule_count = 0  # counts valid green-fill greaterThanOrEqual rules
        range_match_count = 0 # counts CF ranges that cover E2:E26

        # Look through all CF rules for the expected patterns
        for cf_range, rules in cf_rules._cf_rules.items():
            # Check if the range covers E2:E26 (or similar coverage for column E data rows)
            range_str = str(cf_range)
            if 'E2' in range_str and 'E26' in range_str:
                range_match_count += 1

            for rule in rules:
                if rule.type == 'cellIs':
                    # Check for red fill (FFFF0000) with lessThan operator
                    if (rule.operator in ('lessThan', 'lessthan', 'LessThan') and
                            rule.formula and '1' in rule.formula):
                        if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                            fill_color = rule.dxf.fill.fgColor.rgb if rule.dxf.fill.fgColor else ''
                            # Accept red variants: FFFF0000 or similar
                            if fill_color.upper().endswith('FF0000') or 'FF0000' in fill_color.upper():
                                red_rule_count += 1

                    # Check for green fill (FF00FF00) with greaterThanOrEqual operator
                    if (rule.operator in ('greaterThanOrEqual', 'greaterthanorequal', 'GreaterThanOrEqual') and
                            rule.formula and '1' in rule.formula):
                        if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                            fill_color = rule.dxf.fill.fgColor.rgb if rule.dxf.fill.fgColor else ''
                            # Accept green variants: FF00FF00 or similar
                            if fill_color.upper().endswith('00FF00') or '00FF00' in fill_color.upper():
                                green_rule_count += 1

        if red_rule_count >= 1 and green_rule_count >= 1 and range_match_count >= 1:
            print(f"PASS: Component 3 — Conditional formatting on E2:E26 with red (<1) and green (>=1) rules ({0.25} pts)")
            total_score += 0.25
        elif red_rule_count >= 1 or green_rule_count >= 1:
            partial = 0.125
            rules_found = []
            if red_rule_count >= 1:
                rules_found.append('red (<1)')
            if green_rule_count >= 1:
                rules_found.append('green (>=1)')
            print(f"PARTIAL: Component 3 — Found {', '.join(rules_found)} CF rule(s) but missing others ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No conditional formatting found on column E (expected red for <1, green for >=1)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
