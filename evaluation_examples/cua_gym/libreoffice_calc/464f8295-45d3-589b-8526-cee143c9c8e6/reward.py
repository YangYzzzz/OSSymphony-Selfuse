"""
Reward Script: Apply percentage number format to D2:D30
Task ID: calc_gg5_005
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): All 29 cells in D2:D30 have a percentage number format
  Component 2 (0.3): The percentage format shows exactly 1 decimal place (0.0%)
  Component 3 (0.3): Underlying values remain unchanged (still decimals 0-1)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_005'


def persist_app_state(domain):
    """Save any unsaved LibreOffice edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify that D2:D30 on Semester1 sheet have percentage format with 1 decimal.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check Semester1 sheet exists
    if 'Semester1' not in wb.sheetnames:
        print("CRITICAL: 'Semester1' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Semester1']

    # Component 1: All 29 cells D2:D30 have a percentage number format (0.4 points)
    # A percentage format contains '%' character.
    try:
        pct_count = 0
        for row in range(2, 31):
            fmt = ws.cell(row=row, column=4).number_format
            if fmt and '%' in str(fmt):
                pct_count += 1

        if pct_count == 29:
            print(f"PASS: Component 1 — All 29 cells have percentage format ({pct_count}/29) (0.4 pts)")
            total_score += 0.4
        elif pct_count > 0:
            partial = 0.4 * (pct_count / 29)
            print(f"PARTIAL: Component 1 — {pct_count}/29 cells have percentage format ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No cells have percentage format (0/29)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Format is specifically "0.0%" (one decimal place) (0.3 points)
    # Check that the format string indicates exactly 1 decimal place percentage.
    try:
        exact_fmt_count = 0
        for row in range(2, 31):
            fmt = ws.cell(row=row, column=4).number_format
            # Accept common 1-decimal percentage formats
            if fmt in ('0.0%', '0.0 %'):
                exact_fmt_count += 1

        if exact_fmt_count == 29:
            print(f"PASS: Component 2 — All 29 cells have exact 0.0% format ({exact_fmt_count}/29) (0.3 pts)")
            total_score += 0.3
        elif exact_fmt_count > 0:
            partial = 0.3 * (exact_fmt_count / 29)
            print(f"PARTIAL: Component 2 — {exact_fmt_count}/29 cells have exact 0.0% format ({partial:.2f} pts)")
            total_score += partial
        else:
            # Check if they have some other percentage format (partial credit)
            print(f"FAIL: Component 2 — No cells have exact 0.0% format. Sample D2 format: {ws.cell(row=2, column=4).number_format}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Values remain as decimals [0,1] AND have percentage format applied (0.3 points)
    # This is a compound check: the format must be percentage AND the underlying values
    # must still be decimals (not multiplied by 100). Both conditions must hold.
    # This ensures we only score when the task change (format) was applied correctly
    # without corrupting the data.
    try:
        compound_pass_count = 0
        for row in range(2, 31):
            cell = ws.cell(row=row, column=4)
            val = cell.value
            fmt = cell.number_format
            has_pct_fmt = fmt is not None and '%' in str(fmt)
            val_intact = val is not None and isinstance(val, (int, float)) and 0 <= float(val) <= 1
            if has_pct_fmt and val_intact:
                compound_pass_count += 1

        if compound_pass_count == 29:
            print(f"PASS: Component 3 — All 29 cells have pct format + intact values ({compound_pass_count}/29) (0.3 pts)")
            total_score += 0.3
        elif compound_pass_count > 0:
            partial = 0.3 * (compound_pass_count / 29)
            print(f"PARTIAL: Component 3 — {compound_pass_count}/29 cells pass compound check ({partial:.2f} pts)")
            total_score += partial
        else:
            sample_val = ws.cell(row=2, column=4).value
            sample_fmt = ws.cell(row=2, column=4).number_format
            print(f"FAIL: Component 3 — Compound check failed. D2 value={sample_val}, format={sample_fmt}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
