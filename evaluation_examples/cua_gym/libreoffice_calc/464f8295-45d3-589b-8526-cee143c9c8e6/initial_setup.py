"""
Initial Setup: Apply custom percentage format to grade ratios
Task ID: calc_gg5_005
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_005'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Semester1 ---
    ws = wb.active
    ws.title = 'Semester1'

    # Headers
    headers = ['Student ID', 'Name', 'Score', 'Grade Ratio']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 29 rows of realistic student data (D2:D30)
    students = [
        ['S2024001', 'Sarah Chen', 92, 0.875],
        ['S2024002', 'Marcus Johnson', 88, 0.823],
        ['S2024003', 'Aisha Patel', 95, 0.912],
        ['S2024004', 'James O\'Brien', 76, 0.734],
        ['S2024005', 'Maria Garcia', 81, 0.789],
        ['S2024006', 'Wei Zhang', 90, 0.856],
        ['S2024007', 'Emma Thompson', 67, 0.645],
        ['S2024008', 'Carlos Rivera', 84, 0.807],
        ['S2024009', 'Priya Sharma', 91, 0.871],
        ['S2024010', 'David Kim', 73, 0.702],
        ['S2024011', 'Olivia Brown', 86, 0.831],
        ['S2024012', 'Ahmed Hassan', 79, 0.762],
        ['S2024013', 'Sophie Martin', 94, 0.905],
        ['S2024014', 'Tyler Davis', 71, 0.688],
        ['S2024015', 'Yuki Tanaka', 87, 0.839],
        ['S2024016', 'Rachel Green', 82, 0.793],
        ['S2024017', 'Liam Wilson', 69, 0.667],
        ['S2024018', 'Fatima Al-Rashid', 93, 0.898],
        ['S2024019', 'Nathan Scott', 77, 0.745],
        ['S2024020', 'Isabella Rossi', 85, 0.818],
        ['S2024021', 'Ethan Park', 90, 0.862],
        ['S2024022', 'Hannah Lee', 74, 0.716],
        ['S2024023', 'Omar Diallo', 88, 0.847],
        ['S2024024', 'Chloe Dubois', 80, 0.775],
        ['S2024025', 'Ryan Murphy', 96, 0.923],
        ['S2024026', 'Zara Okafor', 83, 0.801],
        ['S2024027', 'Lucas Fernandez', 72, 0.694],
        ['S2024028', 'Mia Anderson', 89, 0.854],
        ['S2024029', 'Daniel Nguyen', 78, 0.751],
    ]

    for r, row_data in enumerate(students, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Ensure D column has General format (no percentage formatting)
    for r in range(2, 31):
        ws.cell(row=r, column=4).number_format = 'General'

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
