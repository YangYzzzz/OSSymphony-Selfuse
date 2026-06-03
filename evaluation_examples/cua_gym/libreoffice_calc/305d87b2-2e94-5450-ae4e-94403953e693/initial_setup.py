"""
Initial Setup: Split the combined product label column into individual columns for Product Name, Brand, and Unit Size.
Task ID: osworld_calc_text_split_columns_009
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_text_split_columns_009'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Grocery Inventory ---
    ws = wb.active
    ws.title = 'Grocery Inventory'

    # Headers
    headers = ['Product Label', 'Brand', 'Product Name', 'Unit Size']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Realistic grocery data — combined labels only; B/C/D left empty (task is to fill via formulas)
    data = [
        'Heinz - Tomato Ketchup (500ml)',
        'Nestle - Condensed Milk (397g)',
        'Kellogs - Corn Flakes (750g)',
        'Unilever - Sunflower Oil (1L)',
        'Del Monte - Diced Tomatoes (400g)',
        "Campbell's - Chicken Soup (305ml)",
        'Quaker - Rolled Oats (1kg)',
        'Hellmann - Mayonnaise (430ml)',
        'Birds Eye - Garden Peas (450g)',
        'Baxters - Lentil Soup (400g)',
        'Princes - Tuna in Brine (185g)',
        'Lee Kum Kee - Soy Sauce (150ml)',
    ]

    for r, label in enumerate(data, 2):
        ws.cell(row=r, column=1, value=label)
        # Columns B, C, D intentionally left empty — agent must enter formulas

    # Set column widths for readability
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 12

    # Freeze the header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
