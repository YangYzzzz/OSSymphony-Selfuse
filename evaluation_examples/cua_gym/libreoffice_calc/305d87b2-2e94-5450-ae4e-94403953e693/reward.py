"""
Reward Script: Split combined product label column into individual columns
Task ID: osworld_calc_text_split_columns_009
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): Brand column (B2:B13) contains formulas extracting brand (before ' - ')
  Component 2 (0.35): Product Name column (C2:C13) contains formulas extracting product name
  Component 3 (0.30): Unit Size column (D2:D13) contains formulas extracting unit size
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_text_split_columns_009'

DATA_ROWS = list(range(2, 14))  # rows 2 through 13 (12 data rows)


def is_brand_formula(value):
    """
    Check if a cell value is a formula that extracts the brand (text before ' - ').
    Expected pattern: =LEFT(A<n>, FIND(" - ", A<n>) - 1)
    Also accept alternative approaches that use LEFT/FIND or equivalent logic.
    """
    if not isinstance(value, str):
        return False
    v = value.upper().replace(" ", "")
    # Must be a formula
    if not v.startswith("="):
        return False
    # Should use LEFT to get text before the separator
    # The key pattern: LEFT( referencing A column, combined with FIND(" - ") or similar
    if "LEFT(" in v and ("FIND(" in v or "SEARCH(" in v):
        return True
    # Also accept: extracting via TRIM/LEFT/MID patterns referencing column A
    # Minimal check: formula references A column and uses text-extraction functions
    return False


def is_product_name_formula(value):
    """
    Check if a cell value is a formula that extracts the product name
    (text between ' - ' and ' (').
    Expected pattern: =MID(A<n>, FIND(" - ",A<n>)+3, FIND(" (",A<n>)-FIND(" - ",A<n>)-3)
    """
    if not isinstance(value, str):
        return False
    v = value.upper().replace(" ", "")
    if not v.startswith("="):
        return False
    # Must use MID with FIND to extract middle portion
    if "MID(" in v and ("FIND(" in v or "SEARCH(" in v):
        # Should reference " - " or similar separator AND "(" bracket
        if '"-"' in v or '" - "' in v.replace(" ", "") or "FIND(" in v:
            return True
    return False


def is_unit_size_formula(value):
    """
    Check if a cell value is a formula that extracts the unit size (text between '(' and ')').
    Expected pattern: =MID(A<n>, FIND("(",A<n>)+1, FIND(")",A<n>)-FIND("(",A<n>)-1)
    """
    if not isinstance(value, str):
        return False
    v = value.upper().replace(" ", "")
    if not v.startswith("="):
        return False
    # Must use MID with FIND for parentheses
    if "MID(" in v and ("FIND(" in v or "SEARCH(" in v):
        # Should reference "(" character
        if '"("' in v or '")"' in v:
            return True
    return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook — fail fast if file is unreadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Access the Grocery Inventory sheet
    try:
        if 'Grocery Inventory' in wb.sheetnames:
            ws = wb['Grocery Inventory']
        else:
            ws = wb.active
    except Exception as e:
        print(f"CRITICAL: Cannot access worksheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: column A (Product Label) must still have data (sanity check)
    try:
        label_a2 = ws['A2'].value
        if not label_a2 or ' - ' not in str(label_a2):
            print("CRITICAL: Column A (Product Label) has been altered or is missing. Cannot score.")
            print("REWARD: 0.0")
            return 0.0
    except Exception as e:
        print(f"CRITICAL: Cannot read A2: {e}")
        print("REWARD: 0.0")
        return 0.0

    # -----------------------------------------------------------------------
    # Component 1: Brand column (B2:B13) has formulas extracting brand (0.35 points)
    # These must FAIL on initial (all None) and PASS on golden (formulas present)
    # -----------------------------------------------------------------------
    try:
        brand_formula_count = 0
        brand_issues = []
        for row in DATA_ROWS:
            cell_val = ws.cell(row=row, column=2).value  # Column B
            if is_brand_formula(cell_val):
                brand_formula_count += 1
            else:
                brand_issues.append(f"B{row}={repr(cell_val)}")

        if brand_formula_count == len(DATA_ROWS):
            print(f"PASS: Component 1 — Brand column B2:B13 all have extraction formulas ({brand_formula_count}/{len(DATA_ROWS)}) (0.35 pts)")
            total_score += 0.35
        elif brand_formula_count > 0:
            partial = round(0.35 * brand_formula_count / len(DATA_ROWS), 4)
            print(f"PARTIAL: Component 1 — Brand column has {brand_formula_count}/{len(DATA_ROWS)} formulas; missing: {brand_issues[:3]}")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Brand column B2:B13 has no extraction formulas (all empty/None)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Product Name column (C2:C13) has formulas extracting product name (0.35 points)
    # -----------------------------------------------------------------------
    try:
        pname_formula_count = 0
        pname_issues = []
        for row in DATA_ROWS:
            cell_val = ws.cell(row=row, column=3).value  # Column C
            if is_product_name_formula(cell_val):
                pname_formula_count += 1
            else:
                pname_issues.append(f"C{row}={repr(cell_val)}")

        if pname_formula_count == len(DATA_ROWS):
            print(f"PASS: Component 2 — Product Name column C2:C13 all have extraction formulas ({pname_formula_count}/{len(DATA_ROWS)}) (0.35 pts)")
            total_score += 0.35
        elif pname_formula_count > 0:
            partial = round(0.35 * pname_formula_count / len(DATA_ROWS), 4)
            print(f"PARTIAL: Component 2 — Product Name column has {pname_formula_count}/{len(DATA_ROWS)} formulas; missing: {pname_issues[:3]}")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Product Name column C2:C13 has no extraction formulas (all empty/None)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Unit Size column (D2:D13) has formulas extracting unit size (0.30 points)
    # -----------------------------------------------------------------------
    try:
        usize_formula_count = 0
        usize_issues = []
        for row in DATA_ROWS:
            cell_val = ws.cell(row=row, column=4).value  # Column D
            if is_unit_size_formula(cell_val):
                usize_formula_count += 1
            else:
                usize_issues.append(f"D{row}={repr(cell_val)}")

        if usize_formula_count == len(DATA_ROWS):
            print(f"PASS: Component 3 — Unit Size column D2:D13 all have extraction formulas ({usize_formula_count}/{len(DATA_ROWS)}) (0.30 pts)")
            total_score += 0.30
        elif usize_formula_count > 0:
            partial = round(0.30 * usize_formula_count / len(DATA_ROWS), 4)
            print(f"PARTIAL: Component 3 — Unit Size column has {usize_formula_count}/{len(DATA_ROWS)} formulas; missing: {usize_issues[:3]}")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Unit Size column D2:D13 has no extraction formulas (all empty/None)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the VM
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
