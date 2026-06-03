"""
Reward Script: Build compensation benchmarking table with gap formulas
Task ID: calc_hr_044
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): E column formulas (Gap vs P50) in E2:E5
  Component 2 (0.35): F column formulas (Gap vs P75) in F2:F5
  Component 3 (0.30): Percentage number format on E2:F5
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_044'


def normalize_formula(f):
    """Normalize formula for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def is_gap_formula(formula, row, ref_col):
    """
    Check if formula computes (B-<ref_col>)/<ref_col> for the given row.
    Accepts variants like =(B2-C2)/C2 or =(B2-C2)/C2 with different spacing.
    ref_col is 'C' for P50 gap, 'D' for P75 gap.
    """
    norm = normalize_formula(formula)
    # Expected pattern: =(B<row>-<ref><row>)/<ref><row>
    expected = f'=(B{row}-{ref_col}{row})/{ref_col}{row}'
    return norm == normalize_formula(expected)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that 'Benchmark' sheet exists (precondition gate)
    if 'Benchmark' not in wb.sheetnames:
        print(f"CRITICAL: 'Benchmark' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Benchmark']

    # Component 1: E column formulas — Gap vs P50 (0.35 points)
    # E2:E5 should contain =(B<row>-C<row>)/C<row>
    try:
        e_pass = 0
        for row in range(2, 6):
            cell = ws.cell(row=row, column=5)  # Column E
            val = cell.value
            if is_gap_formula(val, row, 'C'):
                e_pass += 1
                print(f"PASS: E{row} has correct gap formula: {val}")
            else:
                print(f"FAIL: E{row} expected gap formula =(B{row}-C{row})/C{row}, found: {val!r}")

        if e_pass == 4:
            total_score += 0.35
            print(f"PASS: Component 1 — All E column formulas correct (0.35 pts)")
        elif e_pass > 0:
            partial = round(0.35 * e_pass / 4, 4)
            total_score += partial
            print(f"PARTIAL: Component 1 — {e_pass}/4 E column formulas correct ({partial} pts)")
        else:
            print(f"FAIL: Component 1 — No E column formulas correct (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: F column formulas — Gap vs P75 (0.35 points)
    # F2:F5 should contain =(B<row>-D<row>)/D<row>
    try:
        f_pass = 0
        for row in range(2, 6):
            cell = ws.cell(row=row, column=6)  # Column F
            val = cell.value
            if is_gap_formula(val, row, 'D'):
                f_pass += 1
                print(f"PASS: F{row} has correct gap formula: {val}")
            else:
                print(f"FAIL: F{row} expected gap formula =(B{row}-D{row})/D{row}, found: {val!r}")

        if f_pass == 4:
            total_score += 0.35
            print(f"PASS: Component 2 — All F column formulas correct (0.35 pts)")
        elif f_pass > 0:
            partial = round(0.35 * f_pass / 4, 4)
            total_score += partial
            print(f"PARTIAL: Component 2 — {f_pass}/4 F column formulas correct ({partial} pts)")
        else:
            print(f"FAIL: Component 2 — No F column formulas correct (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Percentage number format on E2:F5 (0.30 points)
    # All 8 cells should have a percentage format (containing '%')
    try:
        pct_pass = 0
        total_cells = 8
        for row in range(2, 6):
            for col in [5, 6]:  # E and F
                cell = ws.cell(row=row, column=col)
                fmt = cell.number_format or 'General'
                coord = cell.coordinate
                if '%' in fmt:
                    pct_pass += 1
                    print(f"PASS: {coord} has percentage format: {fmt}")
                else:
                    print(f"FAIL: {coord} expected percentage format, found: {fmt!r}")

        if pct_pass == total_cells:
            total_score += 0.30
            print(f"PASS: Component 3 — All cells formatted as percentage (0.30 pts)")
        elif pct_pass > 0:
            partial = round(0.30 * pct_pass / total_cells, 4)
            total_score += partial
            print(f"PARTIAL: Component 3 — {pct_pass}/{total_cells} cells with percentage format ({partial} pts)")
        else:
            print(f"FAIL: Component 3 — No cells with percentage format (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
