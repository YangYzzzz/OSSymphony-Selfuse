"""
Initial Setup: Build a compensation benchmarking table with role data and market rates.
Task ID: calc_hr_044
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_044'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Benchmark ---
    ws = wb.active
    ws.title = 'Benchmark'

    # Headers
    headers = ['Role', 'Internal Avg', 'Market P50', 'Market P75', 'Gap vs P50', 'Gap vs P75']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows (realistic compensation data)
    data = [
        ['Software Engineer', 105000, 115000, 135000],
        ['Data Analyst', 78000, 82000, 95000],
        ['Product Manager', 120000, 125000, 148000],
        ['Designer', 88000, 90000, 105000],
    ]

    currency_format = '$#,##0'
    data_align = Alignment(horizontal='center', vertical='center')

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            cell.alignment = data_align
            if c >= 2:  # Currency columns
                cell.number_format = currency_format

    # Columns E and F are intentionally LEFT EMPTY (no formulas, no formatting)
    # The task is for the agent to add formulas and format them as percentages.
    # Add borders to E and F cells so table looks complete structurally
    for r in range(2, 6):
        for c in [5, 6]:
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = data_align

    # Column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
