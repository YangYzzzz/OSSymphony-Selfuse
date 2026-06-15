"""
Initial Setup: Prepare workbook for sharing — hide Internal Notes, rename Draft Report, set tab color
Task ID: calc_sht_multiop_001
Domain: libreoffice_calc

Creates a workbook with four sheets:
  - 'Draft Report': 80 rows of formatted financial data, no tab color
  - 'Internal Notes': 20 rows of reviewer comments, visible
  - 'Supporting Data': 500 rows of source data
  - 'Charts': Visualization sheet
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import datetime

WORKDIR = '/home/user'
TASK_ID = 'calc_sht_multiop_001'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def apply_header_style(cell, bold=True):
    cell.font = Font(name='Calibri', size=11, bold=bold)
    cell.fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='000000')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def create_initial():
    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ #
    # Sheet 1: 'Draft Report' — 80 rows of quarterly financial data
    # ------------------------------------------------------------------ #
    ws_draft = wb.active
    ws_draft.title = 'Draft Report'
    # No tab color (as per context: "Tab currently has no color")

    dr_headers = ['Quarter', 'Region', 'Department', 'Revenue ($)', 'Expenses ($)',
                  'Net Profit ($)', 'YoY Growth (%)', 'Budget Variance ($)', 'Headcount', 'Status']
    for col, h in enumerate(dr_headers, 1):
        cell = ws_draft.cell(row=1, column=col, value=h)
        apply_header_style(cell)

    regions = ['North America', 'Europe', 'Asia Pacific', 'Latin America', 'Middle East']
    departments = ['Engineering', 'Marketing', 'Sales', 'Operations', 'Finance',
                   'HR', 'Legal', 'Product', 'Support', 'Research']
    quarters = ['Q1 2024', 'Q2 2024', 'Q3 2024', 'Q4 2024',
                'Q1 2025', 'Q2 2025', 'Q3 2025', 'Q4 2025']
    statuses = ['On Track', 'Ahead', 'Behind', 'At Risk', 'Completed']

    import random
    random.seed(42)

    dr_data = []
    for i in range(80):
        quarter = quarters[i % len(quarters)]
        region = regions[i % len(regions)]
        dept = departments[i % len(departments)]
        revenue = round(random.uniform(85000, 520000), 2)
        expenses = round(revenue * random.uniform(0.55, 0.82), 2)
        net_profit = round(revenue - expenses, 2)
        yoy = round(random.uniform(-5.5, 28.3), 2)
        budget_var = round(random.uniform(-12000, 35000), 2)
        headcount = random.randint(12, 95)
        status = statuses[i % len(statuses)]
        dr_data.append([quarter, region, dept, revenue, expenses, net_profit, yoy, budget_var, headcount, status])

    for r, row_data in enumerate(dr_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_draft.cell(row=r, column=c, value=val)

    # Column widths
    col_widths = [10, 16, 14, 14, 14, 14, 14, 16, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws_draft.column_dimensions[get_column_letter(i)].width = w

    ws_draft.freeze_panes = 'A2'

    # ------------------------------------------------------------------ #
    # Sheet 2: 'Internal Notes' — 20 rows of reviewer comments, VISIBLE
    # ------------------------------------------------------------------ #
    ws_notes = wb.create_sheet('Internal Notes')
    # Must be visible in initial state

    notes_headers = ['Date', 'Reviewer', 'Section', 'Comment', 'Action Required', 'Priority', 'Status']
    for col, h in enumerate(notes_headers, 1):
        cell = ws_notes.cell(row=1, column=col, value=h)
        apply_header_style(cell)

    reviewers = ['Diana Marsh', 'Kevin Okafor', 'Priya Nair', 'Tom Whitfield', 'Lucia Fernandez']
    sections = ['Revenue Analysis', 'Expense Breakdown', 'Headcount Summary',
                'Forecast', 'Risk Assessment', 'YoY Comparison', 'Regional Summary']
    comments = [
        'Verify Q3 revenue figures against source data before publishing',
        'Expense categories need further breakdown — merge with Supporting Data',
        'Headcount numbers for APAC appear inflated by 12%; recheck with HR',
        'Forecast assumptions for H2 2025 must be updated post board meeting',
        'Risk section references outdated market indices — update to Q1 2025',
        'YoY growth calculation for Latin America uses inconsistent base year',
        'Regional summary totals do not reconcile with Department totals — delta $4,320',
        'Please confirm budget variance sign convention (positive = over/under?)',
        'Chart labels on page 4 are too small for print format',
        'Compliance sign-off still pending for Middle East figures',
        'Add footnote for reclassified R&D expenses from Q4 2024',
        'Marketing spend Q1 2025: need CMO approval before disclosing externally',
        'Engineering headcount: includes 15 contractors — flag clearly',
        'Net profit margin for Finance dept seems unusually high — double-check',
        'Operations Q2 data uses preliminary estimate, flag as subject to revision',
        'Update document header with new fiscal year end date',
        'Legal review of forward-looking statements needed by Friday',
        'HR section: anonymize individual salary data per data protection policy',
        'Support department expenses omit IT allocation — include in final',
        'Final approval workflow: send to CFO then CEO before distribution',
    ]
    priorities = ['High', 'Medium', 'Low', 'Critical', 'Medium']
    note_statuses = ['Open', 'In Progress', 'Resolved', 'Pending Review', 'Open']

    for i in range(20):
        date_val = datetime.date(2025, 1, 15) + datetime.timedelta(days=i * 3)
        ws_notes.cell(row=i+2, column=1, value=str(date_val))
        ws_notes.cell(row=i+2, column=2, value=reviewers[i % len(reviewers)])
        ws_notes.cell(row=i+2, column=3, value=sections[i % len(sections)])
        ws_notes.cell(row=i+2, column=4, value=comments[i])
        ws_notes.cell(row=i+2, column=5, value='Yes' if i % 3 != 2 else 'No')
        ws_notes.cell(row=i+2, column=6, value=priorities[i % len(priorities)])
        ws_notes.cell(row=i+2, column=7, value=note_statuses[i % len(note_statuses)])

    note_widths = [12, 18, 20, 58, 16, 10, 15]
    for i, w in enumerate(note_widths, 1):
        ws_notes.column_dimensions[get_column_letter(i)].width = w

    # ------------------------------------------------------------------ #
    # Sheet 3: 'Supporting Data' — 500 rows of source transaction data
    # ------------------------------------------------------------------ #
    ws_support = wb.create_sheet('Supporting Data')

    sd_headers = ['Transaction ID', 'Date', 'Business Unit', 'Account Code',
                  'Description', 'Debit ($)', 'Credit ($)', 'Currency', 'Approved By']
    for col, h in enumerate(sd_headers, 1):
        cell = ws_support.cell(row=1, column=col, value=h)
        apply_header_style(cell)

    business_units = ['Engineering', 'Marketing', 'Sales', 'Operations', 'Finance',
                      'HR', 'Legal', 'Product']
    account_codes = ['5001', '5002', '5100', '5200', '6001', '6100', '7001', '7200', '8001', '9000']
    descriptions = [
        'Software license renewal', 'Cloud infrastructure costs', 'Conference sponsorship',
        'Travel and expenses', 'Contractor invoices', 'Office supplies',
        'Recruiting fees', 'Legal retainer', 'R&D materials', 'Equipment lease',
        'Marketing campaign spend', 'Training program costs', 'Insurance premium',
        'Utilities', 'Maintenance contract',
    ]
    approvers = ['Sarah Chen', 'Marcus Johnson', 'Aisha Patel', 'Robert Kim',
                 'Elena Vasquez', 'James Obi', 'Natasha Brennan']
    currencies = ['USD', 'EUR', 'GBP', 'JPY', 'AUD', 'CAD']

    random.seed(7)
    for i in range(500):
        txn_id = f'TXN-{2024100 + i:07d}'
        date_val = datetime.date(2024, 1, 1) + datetime.timedelta(days=i % 365)
        bu = business_units[i % len(business_units)]
        acc = account_codes[i % len(account_codes)]
        desc = descriptions[i % len(descriptions)]
        debit = round(random.uniform(500, 45000), 2) if i % 3 != 2 else 0
        credit = round(random.uniform(500, 45000), 2) if i % 3 == 2 else 0
        currency = currencies[i % len(currencies)]
        approver = approvers[i % len(approvers)]
        ws_support.cell(row=i+2, column=1, value=txn_id)
        ws_support.cell(row=i+2, column=2, value=str(date_val))
        ws_support.cell(row=i+2, column=3, value=bu)
        ws_support.cell(row=i+2, column=4, value=acc)
        ws_support.cell(row=i+2, column=5, value=desc)
        ws_support.cell(row=i+2, column=6, value=debit)
        ws_support.cell(row=i+2, column=7, value=credit)
        ws_support.cell(row=i+2, column=8, value=currency)
        ws_support.cell(row=i+2, column=9, value=approver)

    sd_widths = [18, 12, 15, 13, 30, 14, 14, 10, 18]
    for i, w in enumerate(sd_widths, 1):
        ws_support.column_dimensions[get_column_letter(i)].width = w

    ws_support.freeze_panes = 'A2'

    # ------------------------------------------------------------------ #
    # Sheet 4: 'Charts' — Visualization sheet with summary totals
    # ------------------------------------------------------------------ #
    ws_charts = wb.create_sheet('Charts')

    ch_headers = ['Region', 'Q1 2025 Revenue ($)', 'Q2 2025 Revenue ($)',
                  'Q3 2025 Revenue ($)', 'Q4 2025 Revenue ($)', 'FY2025 Total ($)']
    for col, h in enumerate(ch_headers, 1):
        cell = ws_charts.cell(row=1, column=col, value=h)
        apply_header_style(cell)

    chart_data = [
        ['North America', 1_245_300, 1_387_500, 1_452_800, 1_521_000, 5_606_600],
        ['Europe',         892_400,   945_200,   988_700,   1_023_500, 3_849_800],
        ['Asia Pacific',   738_900,   812_300,   895_600,   962_100, 3_408_900],
        ['Latin America',  312_500,   345_800,   378_200,   401_600, 1_438_100],
        ['Middle East',    185_700,   201_300,   219_400,   237_800,   844_200],
    ]
    for r, row_data in enumerate(chart_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_charts.cell(row=r, column=c, value=val)

    # Add a bar chart
    from openpyxl.chart import BarChart, Reference
    chart = BarChart()
    chart.type = 'col'
    chart.title = 'FY2025 Revenue by Region'
    chart.y_axis.title = 'Revenue ($)'
    chart.x_axis.title = 'Region'
    chart.style = 10
    data_ref = Reference(ws_charts, min_col=2, min_row=1, max_col=5, max_row=6)
    cats = Reference(ws_charts, min_col=1, min_row=2, max_row=6)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 20
    chart.height = 14
    ws_charts.add_chart(chart, 'H2')

    ch_widths = [16, 22, 22, 22, 22, 18]
    for i, w in enumerate(ch_widths, 1):
        ws_charts.column_dimensions[get_column_letter(i)].width = w

    # ------------------------------------------------------------------ #
    # Verify sheet order and states
    # ------------------------------------------------------------------ #
    # All four sheets must be visible
    for sheet in wb.worksheets:
        sheet.sheet_state = 'visible'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheets: {wb.sheetnames}')
    for ws in wb.worksheets:
        print(f'  {ws.title}: state={ws.sheet_state}, tab_color={ws.sheet_properties.tabColor}')


create_initial()
