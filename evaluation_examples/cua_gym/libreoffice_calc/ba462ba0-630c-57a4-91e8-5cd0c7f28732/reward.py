"""
Reward Script: Prepare workbook for sharing — hide sheet, rename sheet, set tab color
Task ID: calc_sht_multiop_001
Domain: libreoffice_calc

Task: Hide the 'Internal Notes' sheet, rename 'Draft Report' to 'Final Report',
      and set the 'Final Report' tab color to #70AD47 (green).

Scoring Rubric:
  Component 1: 'Internal Notes' sheet is hidden (0.35 pts)
  Component 2: 'Draft Report' renamed to 'Final Report' (0.35 pts)
  Component 3: 'Final Report' tab color is #70AD47 / ARGB FF70AD47 (0.30 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_sht_multiop_001'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition gate: load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: 'Internal Notes' sheet is hidden (0.35 points)
    # This FAILS on initial (visible) → PASSES on golden (hidden)
    try:
        if 'Internal Notes' in wb.sheetnames:
            ws_notes = wb['Internal Notes']
            state = ws_notes.sheet_state
            if state == 'hidden' or state == 'veryHidden':
                print(f"PASS: Component 1 — 'Internal Notes' sheet is hidden (state={state}) (0.35 pts)")
                total_score += 0.35
            else:
                print(f"FAIL: Component 1 — 'Internal Notes' sheet is NOT hidden (state={state}), expected 'hidden'")
        else:
            print("FAIL: Component 1 — 'Internal Notes' sheet not found in workbook")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Sheet formerly named 'Draft Report' is now renamed to 'Final Report' (0.35 points)
    # This FAILS on initial ('Draft Report' exists, 'Final Report' does not) → PASSES on golden ('Final Report' exists)
    try:
        has_final = 'Final Report' in wb.sheetnames
        has_draft = 'Draft Report' in wb.sheetnames
        if has_final and not has_draft:
            print(f"PASS: Component 2 — Sheet renamed from 'Draft Report' to 'Final Report' (0.35 pts)")
            total_score += 0.35
        elif not has_final and has_draft:
            print(f"FAIL: Component 2 — Sheet still named 'Draft Report', not yet renamed to 'Final Report'")
        elif has_final and has_draft:
            print(f"FAIL: Component 2 — Both 'Draft Report' and 'Final Report' exist; old sheet was not replaced")
        else:
            print(f"FAIL: Component 2 — Neither 'Draft Report' nor 'Final Report' found in workbook (sheets: {wb.sheetnames})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: 'Final Report' tab color is #70AD47 (ARGB FF70AD47) (0.30 points)
    # This FAILS on initial (no tab color on 'Draft Report') → PASSES on golden ('Final Report' has FF70AD47)
    try:
        if 'Final Report' in wb.sheetnames:
            ws_final = wb['Final Report']
            tab_color = ws_final.sheet_properties.tabColor
            if tab_color is not None:
                actual_rgb = str(tab_color.rgb).upper()
                expected_rgb = 'FF70AD47'
                if actual_rgb == expected_rgb:
                    print(f"PASS: Component 3 — 'Final Report' tab color is {actual_rgb} (#70AD47) (0.30 pts)")
                    total_score += 0.30
                else:
                    print(f"FAIL: Component 3 — 'Final Report' tab color is {actual_rgb}, expected {expected_rgb}")
            else:
                print(f"FAIL: Component 3 — 'Final Report' has no tab color set (expected FF70AD47)")
        else:
            print(f"FAIL: Component 3 — 'Final Report' sheet not found, cannot check tab color")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
