"""
Reward Script: Fill total row and total column with SUBTOTAL formulas
Task ID: osworld_calc_fill_totals_008
Domain: libreoffice_calc

Task: Fill the total row and total column in the inventory table, and use
SUBTOTAL instead of SUM so that the totals dynamically respond to filtering.

Structure:
  - Sheet: 'Inventory'
  - Header row: Row 1 (Product, Jan-Dec columns B:M, Annual Total in N)
  - Data rows: Rows 2-13 (12 products)
  - Total row: Row 14 (label 'Total' in column A)
  - Annual Total column: Column N

Scoring:
  Component 1 (0.4): Annual Total column (N2:N13) filled with SUBTOTAL(9, B:M row)
  Component 2 (0.4): Total row (B14:M14) filled with SUBTOTAL(9, column range)
  Component 3 (0.2): Total-of-totals cell (N14) has SUBTOTAL(9, N2:N13) formula
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_fill_totals_008'

# Number of data rows (rows 2..13)
DATA_ROW_START = 2
DATA_ROW_END = 13
TOTAL_ROW = 14

# Column indices
COL_A = 1     # Product label
COL_B = 2     # Jan (first month)
COL_M = 13    # Dec (last month)
COL_N = 14    # Annual Total


def normalize_formula(formula_str):
    """Normalize a formula string for comparison: uppercase, no spaces."""
    if not isinstance(formula_str, str):
        return ''
    return formula_str.upper().replace(' ', '').lstrip('=')


def is_subtotal9_formula(formula_str):
    """Check if cell value is a SUBTOTAL(9,...) formula (not SUM)."""
    if not isinstance(formula_str, str):
        return False
    norm = normalize_formula(formula_str)
    return norm.startswith('SUBTOTAL(9,') or norm.startswith('SUBTOTAL(9,')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Inventory' not in wb.sheetnames:
        print(f"CRITICAL: 'Inventory' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Inventory']

    # -------------------------------------------------------------------------
    # Component 1: Annual Total column (N2:N13) — SUBTOTAL(9, B{row}:M{row})
    # Each of the 12 product rows must have a SUBTOTAL(9, ...) formula in col N
    # (0.4 points — 4 checks of 0.1 each, but awarded all-or-nothing per group)
    # Implementation: award partial credit per cell, max 0.4
    # -------------------------------------------------------------------------
    # Component 1 (0.4 pts): Annual Total column uses SUBTOTAL(9, ...)
    try:
        annual_total_pass = 0
        annual_total_fail = []
        for row in range(DATA_ROW_START, DATA_ROW_END + 1):
            cell = ws.cell(row=row, column=COL_N)
            val = cell.value
            if is_subtotal9_formula(val):
                # Also verify it references the correct row (B{row}:M{row})
                norm = normalize_formula(val)
                expected_range = f'B{row}:M{row}'
                if expected_range.upper() in norm:
                    annual_total_pass += 1
                else:
                    annual_total_fail.append(f'N{row}: formula {val!r} has wrong range (expected B{row}:M{row})')
            else:
                annual_total_fail.append(f'N{row}: not a SUBTOTAL(9,...) formula, found: {val!r}')

        if annual_total_pass == 12 and not annual_total_fail:
            print(f"PASS: Component 1 — All 12 Annual Total cells (N2:N13) use SUBTOTAL(9,B{{row}}:M{{row}}) (0.4 pts)")
            total_score += 0.4
        elif annual_total_pass > 0:
            partial = round((annual_total_pass / 12) * 0.4, 4)
            print(f"PARTIAL: Component 1 — {annual_total_pass}/12 Annual Total cells use SUBTOTAL(9,...): +{partial} pts")
            print(f"  Failures: {annual_total_fail[:3]}")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No Annual Total cells use SUBTOTAL(9,...)")
            print(f"  Sample failures: {annual_total_fail[:3]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Total row (B14:M14) — SUBTOTAL(9, {col}2:{col}13) per month
    # Each of the 12 month columns must have SUBTOTAL(9, ...) in row 14
    # (0.4 points)
    # -------------------------------------------------------------------------
    try:
        total_row_pass = 0
        total_row_fail = []
        col_letters = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
        for col_idx, col_letter in enumerate(col_letters, start=COL_B):
            cell = ws.cell(row=TOTAL_ROW, column=col_idx)
            val = cell.value
            if is_subtotal9_formula(val):
                # Verify it references the correct column (col2:col13)
                norm = normalize_formula(val)
                expected_range = f'{col_letter}2:{col_letter}13'
                if expected_range.upper() in norm:
                    total_row_pass += 1
                else:
                    total_row_fail.append(f'{col_letter}14: formula {val!r} has wrong range (expected {col_letter}2:{col_letter}13)')
            else:
                total_row_fail.append(f'{col_letter}14: not a SUBTOTAL(9,...) formula, found: {val!r}')

        if total_row_pass == 12 and not total_row_fail:
            print(f"PASS: Component 2 — All 12 Total row cells (B14:M14) use SUBTOTAL(9,{{col}}2:{{col}}13) (0.4 pts)")
            total_score += 0.4
        elif total_row_pass > 0:
            partial = round((total_row_pass / 12) * 0.4, 4)
            print(f"PARTIAL: Component 2 — {total_row_pass}/12 Total row cells use SUBTOTAL(9,...): +{partial} pts")
            print(f"  Failures: {total_row_fail[:3]}")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No Total row cells (B14:M14) use SUBTOTAL(9,...)")
            print(f"  Sample failures: {total_row_fail[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Total-of-totals cell N14 — SUBTOTAL(9, N2:N13)
    # The intersection of total row and annual total column must also have
    # a SUBTOTAL formula summing the annual totals column
    # (0.2 points)
    # -------------------------------------------------------------------------
    try:
        cell_n14 = ws.cell(row=TOTAL_ROW, column=COL_N)
        val_n14 = cell_n14.value
        if is_subtotal9_formula(val_n14):
            norm = normalize_formula(val_n14)
            # Accept both N2:N13 and B14:M14 as valid references for this cell
            if 'N2:N13' in norm or 'N2:N13' in norm.upper():
                print(f"PASS: Component 3 — N14 uses SUBTOTAL(9,N2:N13): {val_n14!r} (0.2 pts)")
                total_score += 0.2
            else:
                # Still award points if it's a valid SUBTOTAL formula at N14
                print(f"PARTIAL: Component 3 — N14 has SUBTOTAL(9,...) but unexpected range: {val_n14!r} (0.1 pts)")
                total_score += 0.1
        else:
            print(f"FAIL: Component 3 — N14 is not a SUBTOTAL(9,...) formula, found: {val_n14!r}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the VM environment
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
