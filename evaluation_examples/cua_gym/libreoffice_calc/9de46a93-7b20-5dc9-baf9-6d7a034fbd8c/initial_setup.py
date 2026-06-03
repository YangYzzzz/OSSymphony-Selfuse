"""
Initial Setup: Inventory table with products and monthly stock levels
Task ID: osworld_calc_fill_totals_008
Domain: libreoffice_calc

Creates a spreadsheet with an inventory table where:
- Products are in rows (rows 2-13)
- Monthly stock levels are in columns B-M (Jan-Dec)
- Total row (row 14) and Annual Total column (column N) are EMPTY
  (the agent must fill them with SUBTOTAL formulas)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_fill_totals_008'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Inventory Sheet ---
    ws = wb.active
    ws.title = "Inventory"

    # Style helpers
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2E74B5", end_color="FF2E74B5", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

    total_row_font = Font(name="Calibri", size=11, bold=True)
    total_row_fill = PatternFill(start_color="FFDDEBF7", end_color="FFDDEBF7", fill_type="solid")

    thin = Side(style="thin", color="BFBFBF")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # Column headers: A=Product, B-M=Jan-Dec, N=Annual Total
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    headers = ["Product"] + months + ["Annual Total"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Product data: 12 products with monthly stock levels
    products = [
        "Wireless Keyboard",
        "USB-C Monitor",
        "Ergonomic Mouse",
        "Laptop Stand",
        "HDMI Cable 2m",
        "Webcam HD 1080p",
        "Mechanical Keyboard",
        "USB Hub 7-Port",
        "Noise Cancelling Headset",
        "Portable SSD 1TB",
        "Smart Desk Lamp",
        "Cable Management Kit",
    ]

    # Monthly stock data — realistic variation per product (units in warehouse)
    stock_data = [
        # Jan  Feb  Mar  Apr  May  Jun  Jul  Aug  Sep  Oct  Nov  Dec
        [320, 285, 310, 295, 340, 360, 375, 390, 420, 445, 510, 480],   # Wireless Keyboard
        [145, 130, 155, 160, 175, 180, 165, 190, 210, 225, 260, 240],   # USB-C Monitor
        [280, 265, 290, 310, 325, 345, 355, 370, 395, 415, 480, 460],   # Ergonomic Mouse
        [190, 175, 200, 215, 225, 235, 240, 255, 275, 290, 340, 315],   # Laptop Stand
        [560, 530, 575, 590, 610, 625, 640, 660, 685, 710, 790, 755],   # HDMI Cable 2m
        [210, 195, 220, 230, 245, 255, 265, 280, 300, 315, 365, 345],   # Webcam HD 1080p
        [175, 160, 185, 195, 210, 220, 230, 245, 265, 280, 325, 305],   # Mechanical Keyboard
        [340, 315, 355, 365, 380, 395, 410, 425, 450, 470, 545, 515],   # USB Hub 7-Port
        [265, 245, 275, 285, 300, 310, 320, 340, 360, 380, 440, 415],   # Noise Cancelling Headset
        [130, 120, 140, 150, 160, 165, 170, 185, 200, 215, 250, 235],   # Portable SSD 1TB
        [220, 205, 230, 245, 260, 270, 280, 295, 315, 330, 385, 360],   # Smart Desk Lamp
        [410, 385, 420, 435, 450, 465, 475, 490, 515, 540, 615, 580],   # Cable Management Kit
    ]

    for row_idx, (product, monthly) in enumerate(zip(products, stock_data), 2):
        # Product name (column A)
        name_cell = ws.cell(row=row_idx, column=1, value=product)
        name_cell.font = Font(name="Calibri", size=11)
        name_cell.alignment = left_align
        name_cell.border = thin_border

        # Monthly stock (columns B-M)
        for col_idx, stock in enumerate(monthly, 2):
            cell = ws.cell(row=row_idx, column=col_idx, value=stock)
            cell.font = Font(name="Calibri", size=11)
            cell.alignment = center_align
            cell.border = thin_border

        # Annual Total column N — EMPTY (agent must fill with SUBTOTAL)
        annual_cell = ws.cell(row=row_idx, column=14, value=None)
        annual_cell.border = thin_border

    # Total row (row 14) — label only, all formula cells EMPTY
    total_label = ws.cell(row=14, column=1, value="Total")
    total_label.font = total_row_font
    total_label.fill = total_row_fill
    total_label.alignment = left_align
    total_label.border = thin_border

    # Monthly total cells (columns B-M in row 14) — EMPTY (agent must fill)
    for col_idx in range(2, 14):
        cell = ws.cell(row=14, column=col_idx, value=None)
        cell.font = total_row_font
        cell.fill = total_row_fill
        cell.border = thin_border

    # Grand total cell (N14) — also EMPTY
    grand_cell = ws.cell(row=14, column=14, value=None)
    grand_cell.font = total_row_font
    grand_cell.fill = total_row_fill
    grand_cell.border = thin_border

    # Set column widths for readability
    ws.column_dimensions["A"].width = 26
    for col_letter in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]:
        ws.column_dimensions[col_letter].width = 7
    ws.column_dimensions["N"].width = 14

    # Freeze pane below header row
    ws.freeze_panes = "B2"

    # Auto-filter on data range
    ws.auto_filter.ref = "A1:N13"

    wb.save(OUTPUT)
    print(f"Initial file created: {OUTPUT}")

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print("GUI_READY: launched LibreOffice Calc with DISPLAY=:0")


create_initial()
