"""
Initial Setup: Multi-currency deal tracker with exchange rates lookup
Task ID: calc_sales_074
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_074'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Deals ---
    ws1 = wb.active
    ws1.title = 'Deals'

    # Headers
    headers = ['Deal', 'Currency', 'Local Value', 'Exchange Rate', 'USD Value']
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align

    # Deal data (columns A-C only; D and E left empty for the task)
    deals = [
        ['D1', 'EUR', 85000],
        ['D2', 'GBP', 120000],
        ['D3', 'USD', 200000],
        ['D4', 'JPY', 15000000],
        ['D5', 'EUR', 150000],
        ['D6', 'CAD', 95000],
    ]
    currency_format = '#,##0'
    for r, row_data in enumerate(deals, 2):
        ws1.cell(row=r, column=1, value=row_data[0])
        ws1.cell(row=r, column=2, value=row_data[1])
        c_cell = ws1.cell(row=r, column=3, value=row_data[2])
        c_cell.number_format = currency_format

    # Column widths for readability
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 16
    ws1.column_dimensions['D'].width = 16
    ws1.column_dimensions['E'].width = 18

    # --- Sheet 2: FXRates ---
    ws2 = wb.create_sheet('FXRates')

    fx_headers = ['Currency', 'USD Rate']
    for col, h in enumerate(fx_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FF548235", end_color="FF548235", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    fx_data = [
        ['EUR', 1.08],
        ['GBP', 1.27],
        ['USD', 1.00],
        ['JPY', 0.0067],
        ['CAD', 0.74],
    ]
    for r, row_data in enumerate(fx_data, 2):
        ws2.cell(row=r, column=1, value=row_data[0])
        rate_cell = ws2.cell(row=r, column=2, value=row_data[1])
        rate_cell.number_format = '0.0000'

    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
