"""
Reward Script: Multi-currency deal tracker with VLOOKUP and USD conversion
Task ID: calc_sales_074
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): VLOOKUP formulas in D2:D7 referencing FXRates sheet
  Component 2 (0.35): USD Value formulas in E2:E7 (=C*D multiplication)
  Component 3 (0.30): Total Pipeline label in A8 + SUM formula in E8
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_074'


def persist_app_state(domain):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def normalize_formula(val):
    """Normalize a formula string for comparison: uppercase, strip spaces."""
    if not isinstance(val, str):
        return ""
    return val.upper().replace(" ", "")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that required sheets exist
    if 'Deals' not in wb.sheetnames:
        print("CRITICAL: 'Deals' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Deals']

    # Component 1: VLOOKUP formulas in D2:D7 (0.35 points)
    # Each correct VLOOKUP earns ~0.058 points (0.35/6)
    try:
        vlookup_count = 0
        expected_rows = [2, 3, 4, 5, 6, 7]
        for row in expected_rows:
            cell_val = ws.cell(row=row, column=4).value  # Column D
            if cell_val and isinstance(cell_val, str):
                norm = normalize_formula(cell_val)
                # Accept VLOOKUP referencing B column and FXRates sheet
                if "VLOOKUP" in norm and "FXRATES" in norm:
                    vlookup_count += 1
                else:
                    print(f"FAIL: D{row} has '{cell_val}', expected VLOOKUP referencing FXRates")
            else:
                print(f"FAIL: D{row} is '{cell_val}', expected VLOOKUP formula")

        if vlookup_count == 6:
            print(f"PASS: Component 1 -- All 6 VLOOKUP formulas found in D2:D7 (0.35 pts)")
            total_score += 0.35
        elif vlookup_count > 0:
            partial = round(0.35 * vlookup_count / 6, 4)
            print(f"PARTIAL: Component 1 -- {vlookup_count}/6 VLOOKUP formulas correct ({partial} pts)")
            total_score += partial
        else:
            print("FAIL: Component 1 -- No VLOOKUP formulas found in D2:D7")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: USD Value formulas in E2:E7 (0.35 points)
    # Each cell should have a formula multiplying C column by D column (=C*D)
    try:
        usd_count = 0
        for row in expected_rows:
            cell_val = ws.cell(row=row, column=5).value  # Column E
            if cell_val and isinstance(cell_val, str):
                norm = normalize_formula(cell_val)
                # Accept formulas like =C2*D2 or =D2*C2
                c_ref = f"C{row}"
                d_ref = f"D{row}"
                if (c_ref in norm and d_ref in norm and "*" in norm):
                    usd_count += 1
                else:
                    print(f"FAIL: E{row} has '{cell_val}', expected multiplication of C{row}*D{row}")
            else:
                print(f"FAIL: E{row} is '{cell_val}', expected formula (=C{row}*D{row})")

        if usd_count == 6:
            print(f"PASS: Component 2 -- All 6 USD Value formulas found in E2:E7 (0.35 pts)")
            total_score += 0.35
        elif usd_count > 0:
            partial = round(0.35 * usd_count / 6, 4)
            print(f"PARTIAL: Component 2 -- {usd_count}/6 USD Value formulas correct ({partial} pts)")
            total_score += partial
        else:
            print("FAIL: Component 2 -- No USD Value formulas found in E2:E7")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Total Pipeline label in A8 + SUM formula in E8 (0.30 points)
    try:
        comp3_score = 0.0

        # Sub-check 3a: A8 contains "Total Pipeline (USD)" or similar label (0.10 pts)
        a8_val = ws.cell(row=8, column=1).value
        if a8_val and "total" in str(a8_val).lower() and "pipeline" in str(a8_val).lower():
            print(f"PASS: Component 3a -- A8 label found: '{a8_val}' (0.10 pts)")
            comp3_score += 0.10
        else:
            print(f"FAIL: Component 3a -- A8 is '{a8_val}', expected 'Total Pipeline (USD)'")

        # Sub-check 3b: E8 contains a SUM formula over E2:E7 (0.20 pts)
        e8_val = ws.cell(row=8, column=5).value
        if e8_val and isinstance(e8_val, str):
            norm = normalize_formula(e8_val)
            if "SUM" in norm and "E" in norm:
                print(f"PASS: Component 3b -- E8 SUM formula found: '{e8_val}' (0.20 pts)")
                comp3_score += 0.20
            else:
                print(f"FAIL: Component 3b -- E8 has '{e8_val}', expected SUM formula")
        else:
            print(f"FAIL: Component 3b -- E8 is '{e8_val}', expected SUM formula")

        if comp3_score > 0:
            total_score += comp3_score
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
