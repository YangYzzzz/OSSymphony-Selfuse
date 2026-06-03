"""
Reward Script: Compare current vs. refinanced mortgage payments using PMT formula
Task ID: calc_fmb_pmt_refinance_079
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: B8 contains a PMT formula for the current loan (0.5 points)
               Expected: =PMT(B3/B5,B4*B5,-B2) or semantically equivalent
  Component 2: C8 contains a PMT formula for the refinanced loan (0.5 points)
               Expected: =PMT(C3/C5,C4*C5,-C2) or semantically equivalent
  Total: 1.0

The initial file has B8 and C8 empty (None). Only the golden file has these PMT formulas.
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmb_pmt_refinance_079'


def normalize_pmt_formula(formula_str):
    """
    Normalize a PMT formula string for comparison:
    - Strip leading '='
    - Uppercase
    - Remove all whitespace
    """
    if not isinstance(formula_str, str):
        return ''
    return formula_str.upper().replace(' ', '').strip()


def contains_pmt_function(formula_str):
    """Check if the formula string starts with =PMT( (case-insensitive)."""
    if not isinstance(formula_str, str):
        return False
    normalized = normalize_pmt_formula(formula_str)
    return normalized.startswith('=PMT(')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    The task requires:
    1. Cell B8 must contain a PMT formula for the current loan:
       =PMT(B3/B5, B4*B5, -B2) — current loan: 7.2% for 22 years on $320,000
    2. Cell C8 must contain a PMT formula for the refinanced loan:
       =PMT(C3/C5, C4*C5, -C2) — refinanced loan: 5.8% for 30 years on $320,000
    """
    total_score = 0.0

    # Precondition gate: load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: verify the expected sheet exists
    expected_sheet = 'Refinance Comparison'
    if expected_sheet not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{expected_sheet}' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[expected_sheet]

    # Component 1: B8 contains a valid PMT formula for the current loan (0.5 points)
    # Expected formula: =PMT(B3/B5,B4*B5,-B2)
    # This FAILS on initial file (B8 is empty) and PASSES on golden file
    try:
        b8_value = ws['B8'].value
        b8_normalized = normalize_pmt_formula(b8_value) if isinstance(b8_value, str) else ''

        # Check 1a: The cell must contain a PMT formula
        if not contains_pmt_function(b8_value):
            print(f"FAIL: Component 1 — B8 does not contain a PMT formula. Found: {repr(b8_value)}")
        else:
            # Check 1b: The PMT formula must reference the correct cells for current loan
            # Required references: B3 (rate), B5 (payments/year), B4 (years), B2 (balance)
            # Canonical form: =PMT(B3/B5,B4*B5,-B2)
            expected_b8 = '=PMT(B3/B5,B4*B5,-B2)'
            expected_b8_norm = normalize_pmt_formula(expected_b8)

            if b8_normalized == expected_b8_norm:
                print(f"PASS: Component 1 — B8 contains correct PMT formula: {repr(b8_value)} (0.5 pts)")
                total_score += 0.5
            else:
                # Accept semantically equivalent variants — still using B-column references
                # Must reference B3, B5, B4, B2 for correctness
                refs_correct = all(ref in b8_normalized for ref in ['B3', 'B5', 'B4', 'B2'])
                if refs_correct and contains_pmt_function(b8_value):
                    print(f"PASS: Component 1 — B8 contains PMT formula with correct cell refs: {repr(b8_value)} (0.5 pts)")
                    total_score += 0.5
                else:
                    print(f"FAIL: Component 1 — B8 PMT formula has incorrect cell references.")
                    print(f"  Expected: {expected_b8}, Found: {repr(b8_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not check B8: {e}")

    # Component 2: C8 contains a valid PMT formula for the refinanced loan (0.5 points)
    # Expected formula: =PMT(C3/C5,C4*C5,-C2)
    # This FAILS on initial file (C8 is empty) and PASSES on golden file
    try:
        c8_value = ws['C8'].value
        c8_normalized = normalize_pmt_formula(c8_value) if isinstance(c8_value, str) else ''

        # Check 2a: The cell must contain a PMT formula
        if not contains_pmt_function(c8_value):
            print(f"FAIL: Component 2 — C8 does not contain a PMT formula. Found: {repr(c8_value)}")
        else:
            # Check 2b: The PMT formula must reference the correct cells for refinanced loan
            # Required references: C3 (rate), C5 (payments/year), C4 (years), C2 (balance)
            # Canonical form: =PMT(C3/C5,C4*C5,-C2)
            expected_c8 = '=PMT(C3/C5,C4*C5,-C2)'
            expected_c8_norm = normalize_pmt_formula(expected_c8)

            if c8_normalized == expected_c8_norm:
                print(f"PASS: Component 2 — C8 contains correct PMT formula: {repr(c8_value)} (0.5 pts)")
                total_score += 0.5
            else:
                # Accept semantically equivalent variants — still using C-column references
                # Must reference C3, C5, C4, C2 for correctness
                refs_correct = all(ref in c8_normalized for ref in ['C3', 'C5', 'C4', 'C2'])
                if refs_correct and contains_pmt_function(c8_value):
                    print(f"PASS: Component 2 — C8 contains PMT formula with correct cell refs: {repr(c8_value)} (0.5 pts)")
                    total_score += 0.5
                else:
                    print(f"FAIL: Component 2 — C8 PMT formula has incorrect cell references.")
                    print(f"  Expected: {expected_c8}, Found: {repr(c8_value)}")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not check C8: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
