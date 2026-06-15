"""
Initial Setup: Compare current vs. refinanced mortgage payments
Task ID: calc_fmb_pmt_refinance_079
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_pmt_refinance_079'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Refinance Comparison'

    # --- Column headers in row 1 ---
    ws['A1'] = 'Parameter'
    ws['B1'] = 'Current Loan'
    ws['C1'] = 'Refinanced Loan'

    # Style header row
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF', size=11)
    for col_letter in ['A', 'B', 'C']:
        cell = ws[f'{col_letter}1']
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Row 2: Remaining Balance ---
    ws['A2'] = 'Remaining Balance'
    ws['B2'] = 320000
    ws['C2'] = 320000
    ws['B2'].number_format = '$#,##0.00'
    ws['C2'].number_format = '$#,##0.00'

    # --- Row 3: Annual Rate ---
    ws['A3'] = 'Annual Rate'
    ws['B3'] = 0.072
    ws['C3'] = 0.058
    ws['B3'].number_format = '0.000%'
    ws['C3'].number_format = '0.000%'

    # --- Row 4: Years Remaining ---
    ws['A4'] = 'Years Remaining'
    ws['B4'] = 22
    ws['C4'] = 30

    # --- Row 5: Payments/Year ---
    ws['A5'] = 'Payments/Year'
    ws['B5'] = 12
    ws['C5'] = 12

    # --- Row 6: blank separator ---

    # --- Row 7: Monthly Payment header ---
    ws['A7'] = 'Monthly Payment'
    ws['A7'].font = Font(bold=True)

    # B8 and C8 are intentionally left empty (target cells for PMT formulas)
    # Row 8 has labels
    ws['A8'] = 'Calculated Payment'

    # --- Column widths for readability ---
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18

    # --- Add a note/context in row 10+ ---
    ws['A10'] = 'Loan Details'
    ws['A10'].font = Font(bold=True, italic=True)
    ws['A11'] = 'Lender'
    ws['B11'] = 'First Metro Bank'
    ws['C11'] = 'Pacific Lending Group'
    ws['A12'] = 'Loan Purpose'
    ws['B12'] = 'Primary Residence'
    ws['C12'] = 'Primary Residence'
    ws['A13'] = 'Origination Date'
    ws['B13'] = '2003-05-15'
    ws['C13'] = '2026-01-01'
    ws['A14'] = 'Loan Type'
    ws['B14'] = 'Fixed-Rate Mortgage'
    ws['C14'] = 'Fixed-Rate Mortgage'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheet: Refinance Comparison')
    print('B8 and C8 are empty (ready for PMT formulas)')


create_initial()
