"""
Reward Script: Insert hyperlink in cell D4 of 'Resources' sheet
Task ID: calc_gg3_023
Domain: libreoffice_calc
Scoring:
  Component 1: D4 has a hyperlink object (0.3 pts)
  Component 2: Hyperlink target URL is 'https://docs.libreoffice.org' (0.3 pts)
  Component 3: D4 display text is 'LibreOffice Documentation' (0.4 pts)
"""

import os

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_023'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Resources' sheet must exist
    if 'Resources' not in wb.sheetnames:
        print("FAIL: 'Resources' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Resources']

    # Component 1: D4 has a hyperlink object (0.3 points)
    # Initial state: D4 has no hyperlink. Golden state: D4 has a hyperlink.
    try:
        cell = ws['D4']
        if cell.hyperlink is not None:
            print(f"PASS: Component 1 — D4 has a hyperlink (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — D4 has no hyperlink")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Hyperlink target URL is 'https://docs.libreoffice.org' (0.3 points)
    # Initial state: no hyperlink, so no target. Golden: target is the correct URL.
    try:
        cell = ws['D4']
        if cell.hyperlink is not None and cell.hyperlink.target is not None:
            target = cell.hyperlink.target.rstrip('/')
            expected = 'https://docs.libreoffice.org'
            if target == expected:
                print(f"PASS: Component 2 — Hyperlink target is '{target}' (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — Expected target '{expected}', found '{target}'")
        else:
            print(f"FAIL: Component 2 — No hyperlink or no target on D4")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: D4 display text is 'LibreOffice Documentation' (0.4 points)
    # Initial state: D4 value is 'LibreOffice Help'. Golden: 'LibreOffice Documentation'.
    try:
        cell = ws['D4']
        val = cell.value
        if val is not None and str(val).strip() == 'LibreOffice Documentation':
            print(f"PASS: Component 3 — D4 display text is 'LibreOffice Documentation' (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 3 — Expected 'LibreOffice Documentation', found '{val}'")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
