"""
Initial Setup: Insert hyperlink in cell D4 of Resources sheet
Task ID: calc_gg3_023
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_023'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Resources ---
    ws1 = wb.active
    ws1.title = 'Resources'

    # Headers
    headers = ['Resource Name', 'Category', 'Description', 'Link']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    white_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows - realistic team resource list
    # Columns: Resource Name (A), Category (B), Description (C), Link (D)
    data = [
        ['Python Official Docs', 'Programming', 'Complete Python language reference and tutorials', 'https://docs.python.org'],
        ['Confluence Wiki', 'Internal', 'Team knowledge base and project documentation', 'https://wiki.internal.corp'],
        ['LibreOffice Help', 'Productivity', 'Official documentation for LibreOffice suite', 'LibreOffice Help'],
        ['Jira Board', 'Project Mgmt', 'Sprint tracking and issue management dashboard', 'https://jira.internal.corp'],
        ['GitHub Enterprise', 'Development', 'Source code repositories and CI/CD pipelines', 'https://github.enterprise.corp'],
        ['Slack Workspace', 'Communication', 'Team messaging and channel collaboration', 'https://team.slack.com'],
        ['Grafana Dashboard', 'Monitoring', 'Service health metrics and alerting overview', 'https://grafana.internal.corp'],
        ['AWS Console', 'Infrastructure', 'Cloud resource management and deployment', 'https://console.aws.amazon.com'],
        ['Figma Design', 'Design', 'UI/UX design files and component library', 'https://figma.com/team-project'],
        ['HR Portal', 'Admin', 'Leave requests, payroll, and benefits management', 'https://hr.internal.corp'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.font = Font(name='Calibri', size=11)

    # Set column widths for readability
    ws1.column_dimensions['A'].width = 24
    ws1.column_dimensions['B'].width = 16
    ws1.column_dimensions['C'].width = 50
    ws1.column_dimensions['D'].width = 35

    # NOTE: Cell D4 contains plain text "LibreOffice Help" — NO hyperlink
    # The task is to convert this to a hyperlink

    # --- Sheet 2: Team ---
    ws2 = wb.create_sheet('Team')

    team_headers = ['Name', 'Role', 'Email', 'Department']
    for col, h in enumerate(team_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    team_data = [
        ['Sarah Chen', 'Tech Lead', 'sarah.chen@company.com', 'Engineering'],
        ['Marcus Johnson', 'Product Manager', 'marcus.j@company.com', 'Product'],
        ['Priya Patel', 'UX Designer', 'priya.p@company.com', 'Design'],
        ['David Kim', 'Backend Developer', 'david.k@company.com', 'Engineering'],
        ['Emma Wilson', 'QA Engineer', 'emma.w@company.com', 'Quality'],
        ['Carlos Rivera', 'DevOps Engineer', 'carlos.r@company.com', 'Infrastructure'],
        ['Aisha Mohammed', 'Data Analyst', 'aisha.m@company.com', 'Analytics'],
    ]

    for r, row_data in enumerate(team_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.font = Font(name='Calibri', size=11)

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 30
    ws2.column_dimensions['D'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
