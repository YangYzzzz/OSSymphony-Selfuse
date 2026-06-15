"""
Initial Setup: Add a bold 'Total' summary row at the bottom of the sales data table
Task ID: calc_gsd_001
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_001'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Q4 Data'

    # Default font for the sheet
    base_font = Font(name='Liberation Sans', size=11)

    # Headers in row 1
    headers = ['Salesperson', 'Region', 'Jan', 'Feb', 'Mar', 'Q4 Total']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name='Liberation Sans', size=11, bold=True)

    # 20 salespeople with realistic data
    data = [
        ['Sarah Chen', 'West', 4520, 5130, 4870, 14520],
        ['Marcus Johnson', 'East', 3890, 4210, 4550, 12650],
        ['Priya Patel', 'South', 5200, 4780, 5340, 15320],
        ['James O\'Brien', 'North', 3150, 3420, 3680, 10250],
        ['Aisha Williams', 'West', 6100, 5890, 6230, 18220],
        ['Carlos Rivera', 'East', 4750, 4980, 4620, 14350],
        ['Emily Zhang', 'South', 3680, 3950, 4120, 11750],
        ['David Kim', 'North', 5430, 5100, 5670, 16200],
        ['Rachel Foster', 'West', 4210, 4560, 4380, 13150],
        ['Omar Hassan', 'East', 3970, 4150, 4320, 12440],
        ['Lisa Nakamura', 'South', 5680, 5420, 5890, 16990],
        ['Thomas Wright', 'North', 3340, 3580, 3760, 10680],
        ['Natasha Volkov', 'West', 4890, 5230, 4970, 15090],
        ['Andre Baptiste', 'East', 4120, 4350, 4580, 13050],
        ['Mei-Lin Wu', 'South', 5950, 5710, 6120, 17780],
        ['Kevin Murphy', 'North', 3560, 3820, 3990, 11370],
        ['Fatima Al-Rashid', 'West', 4680, 4910, 4750, 14340],
        ['Ryan Okafor', 'East', 3790, 4080, 4260, 12130],
        ['Sofia Gonzalez', 'South', 5310, 5070, 5490, 15870],
        ['Henrik Larsen', 'North', 4050, 4290, 4430, 12770],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = base_font

    # Set column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14

    # Row 22 is intentionally left empty — the agent must fill it

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
