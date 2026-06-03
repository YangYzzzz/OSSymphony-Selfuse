"""
Reward Script: Add bold 'Total' summary row with SUM formulas, thick bottom border, and light gray background
Task ID: calc_gsd_001
Domain: libreoffice_calc
Scoring:
  Component 1 (0.15): A22 contains 'Total'
  Component 2 (0.30): C22:F22 contain SUM formulas summing rows 2-21
  Component 3 (0.20): A22:F22 are all bold
  Component 4 (0.20): A22:F22 have thick bottom border
  Component 5 (0.15): A22:F22 have light gray background (#D9D9D9)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_001'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Q4 Data' sheet must exist
    if 'Q4 Data' not in wb.sheetnames:
        print(f"CRITICAL: Sheet 'Q4 Data' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Q4 Data']

    # Component 1: A22 contains text 'Total' (0.15 points)
    try:
        val_a22 = ws.cell(row=22, column=1).value
        if val_a22 is not None and str(val_a22).strip().lower() == 'total':
            print(f"PASS: Component 1 — A22 contains 'Total' (value: {val_a22}) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — Expected 'Total' in A22, found: {repr(val_a22)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: C22:F22 contain SUM formulas summing rows 2-21 (0.30 points)
    # Each correct formula earns 0.075 points (4 cells)
    try:
        formula_score = 0.0
        expected_cols = {'C': 3, 'D': 4, 'E': 5, 'F': 6}
        for col_letter, col_idx in expected_cols.items():
            cell_val = ws.cell(row=22, column=col_idx).value
            if cell_val is not None and isinstance(cell_val, str):
                normalized = cell_val.upper().replace(' ', '')
                expected_formula = f'=SUM({col_letter}2:{col_letter}21)'.upper()
                if normalized == expected_formula:
                    formula_score += 0.075
                    print(f"  PASS: {col_letter}22 has correct SUM formula: {cell_val}")
                else:
                    print(f"  FAIL: {col_letter}22 expected {expected_formula}, found: {cell_val}")
            else:
                print(f"  FAIL: {col_letter}22 has no formula, found: {repr(cell_val)}")

        if formula_score > 0:
            print(f"PASS: Component 2 — SUM formulas ({formula_score:.3f} pts)")
            total_score += formula_score
        else:
            print(f"FAIL: Component 2 — No correct SUM formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: A22:F22 are all bold (0.20 points)
    try:
        bold_count = 0
        for col_idx in range(1, 7):
            cell = ws.cell(row=22, column=col_idx)
            if cell.font.bold is True:
                bold_count += 1
            else:
                print(f"  FAIL: Cell row=22, col={col_idx} is not bold (bold={cell.font.bold})")

        if bold_count == 6:
            print(f"PASS: Component 3 — All 6 cells in row 22 are bold (0.20 pts)")
            total_score += 0.20
        elif bold_count > 0:
            partial = 0.20 * (bold_count / 6)
            print(f"PARTIAL: Component 3 — {bold_count}/6 cells bold ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No cells in row 22 are bold")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: A22:F22 have thick bottom border (0.20 points)
    try:
        border_count = 0
        for col_idx in range(1, 7):
            cell = ws.cell(row=22, column=col_idx)
            bottom_style = None
            if cell.border and cell.border.bottom:
                bottom_style = cell.border.bottom.style
            # Accept 'thick' or 'medium' as valid thick border styles
            if bottom_style in ('thick', 'medium'):
                border_count += 1
            else:
                print(f"  FAIL: Cell row=22, col={col_idx} bottom border: {bottom_style}")

        if border_count == 6:
            print(f"PASS: Component 4 — All 6 cells have thick bottom border (0.20 pts)")
            total_score += 0.20
        elif border_count > 0:
            partial = 0.20 * (border_count / 6)
            print(f"PARTIAL: Component 4 — {border_count}/6 cells have thick bottom border ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — No cells in row 22 have thick bottom border")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: A22:F22 have light gray background (#D9D9D9 = FFD9D9D9 in ARGB) (0.15 points)
    try:
        fill_count = 0
        for col_idx in range(1, 7):
            cell = ws.cell(row=22, column=col_idx)
            fill_rgb = None
            try:
                fill_rgb = cell.fill.fgColor.rgb if cell.fill.fgColor else None
            except:
                pass
            fill_type = cell.fill.fill_type

            if fill_type == 'solid' and fill_rgb is not None:
                # Accept FFD9D9D9 or 00D9D9D9 (with or without alpha)
                rgb_hex = str(fill_rgb).upper()
                if rgb_hex.endswith('D9D9D9'):
                    fill_count += 1
                else:
                    print(f"  FAIL: Cell row=22, col={col_idx} fill color: {fill_rgb} (expected *D9D9D9)")
            else:
                print(f"  FAIL: Cell row=22, col={col_idx} fill_type={fill_type}, fill_rgb={fill_rgb}")

        if fill_count == 6:
            print(f"PASS: Component 5 — All 6 cells have light gray background (0.15 pts)")
            total_score += 0.15
        elif fill_count > 0:
            partial = 0.15 * (fill_count / 6)
            print(f"PARTIAL: Component 5 — {fill_count}/6 cells have gray background ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — No cells in row 22 have light gray background")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score:.3f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
