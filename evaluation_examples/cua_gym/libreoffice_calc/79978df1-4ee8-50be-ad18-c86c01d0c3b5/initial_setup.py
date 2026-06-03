"""
Initial Setup: Warehouse operations data with movement records in Sheet1, empty Sheet2.
Task ID: osworld_calc_sheet2_summary_table_003
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_sheet2_summary_table_003'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: Warehouse Movement Records ---
    ws1 = wb.active
    ws1.title = 'Sheet1'

    # Headers
    headers = ['Date', 'Movement Type', 'Product', 'Units']
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Set column widths
    ws1.column_dimensions['A'].width = 14
    ws1.column_dimensions['B'].width = 16
    ws1.column_dimensions['C'].width = 22
    ws1.column_dimensions['D'].width = 10

    # Warehouse movement data - realistic content covering all 12 months of 2024
    # Each entry: (date_string, movement_type, product, units)
    movement_data = [
        # January 2024
        ('2024-01-03', 'Inbound',  'Steel Pipes DN50',        320),
        ('2024-01-05', 'Inbound',  'Hydraulic Fittings M20',  150),
        ('2024-01-08', 'Outbound', 'Steel Pipes DN50',        200),
        ('2024-01-12', 'Inbound',  'Industrial Bearings 6205',480),
        ('2024-01-15', 'Outbound', 'Hydraulic Fittings M20',   90),
        ('2024-01-19', 'Outbound', 'Industrial Bearings 6205', 210),
        ('2024-01-23', 'Inbound',  'Rubber Gaskets 50mm',     600),
        ('2024-01-28', 'Outbound', 'Rubber Gaskets 50mm',     340),
        # February 2024
        ('2024-02-02', 'Inbound',  'Stainless Bolts M12',     800),
        ('2024-02-06', 'Outbound', 'Stainless Bolts M12',     450),
        ('2024-02-09', 'Inbound',  'PVC Tubing 25mm',         270),
        ('2024-02-14', 'Inbound',  'Steel Pipes DN50',        180),
        ('2024-02-18', 'Outbound', 'PVC Tubing 25mm',         130),
        ('2024-02-22', 'Outbound', 'Steel Pipes DN50',        160),
        ('2024-02-26', 'Inbound',  'Industrial Bearings 6205',350),
        # March 2024
        ('2024-03-01', 'Inbound',  'Hydraulic Fittings M20',  220),
        ('2024-03-05', 'Outbound', 'Stainless Bolts M12',     300),
        ('2024-03-08', 'Inbound',  'Rubber Gaskets 50mm',     450),
        ('2024-03-12', 'Outbound', 'Industrial Bearings 6205',190),
        ('2024-03-15', 'Inbound',  'Steel Pipes DN50',        260),
        ('2024-03-19', 'Outbound', 'Hydraulic Fittings M20',  180),
        ('2024-03-22', 'Inbound',  'PVC Tubing 25mm',         310),
        ('2024-03-27', 'Outbound', 'Rubber Gaskets 50mm',     280),
        ('2024-03-29', 'Outbound', 'PVC Tubing 25mm',         200),
        # April 2024
        ('2024-04-02', 'Inbound',  'Stainless Bolts M12',     500),
        ('2024-04-05', 'Inbound',  'Industrial Bearings 6205',420),
        ('2024-04-09', 'Outbound', 'Steel Pipes DN50',        175),
        ('2024-04-12', 'Outbound', 'Stainless Bolts M12',     360),
        ('2024-04-16', 'Inbound',  'Hydraulic Fittings M20',  290),
        ('2024-04-19', 'Outbound', 'Industrial Bearings 6205',250),
        ('2024-04-23', 'Inbound',  'Rubber Gaskets 50mm',     380),
        ('2024-04-26', 'Outbound', 'Hydraulic Fittings M20',  210),
        # May 2024
        ('2024-05-03', 'Inbound',  'Steel Pipes DN50',        295),
        ('2024-05-07', 'Outbound', 'Rubber Gaskets 50mm',     220),
        ('2024-05-10', 'Inbound',  'PVC Tubing 25mm',         400),
        ('2024-05-14', 'Outbound', 'PVC Tubing 25mm',         310),
        ('2024-05-17', 'Inbound',  'Stainless Bolts M12',     600),
        ('2024-05-21', 'Outbound', 'Steel Pipes DN50',        245),
        ('2024-05-24', 'Inbound',  'Industrial Bearings 6205',360),
        ('2024-05-28', 'Outbound', 'Stainless Bolts M12',     420),
        # June 2024
        ('2024-06-04', 'Inbound',  'Hydraulic Fittings M20',  330),
        ('2024-06-07', 'Outbound', 'Industrial Bearings 6205',180),
        ('2024-06-11', 'Inbound',  'Rubber Gaskets 50mm',     510),
        ('2024-06-14', 'Outbound', 'Hydraulic Fittings M20',  260),
        ('2024-06-18', 'Inbound',  'Steel Pipes DN50',        200),
        ('2024-06-21', 'Outbound', 'Rubber Gaskets 50mm',     390),
        ('2024-06-25', 'Inbound',  'PVC Tubing 25mm',         270),
        ('2024-06-28', 'Outbound', 'Steel Pipes DN50',        155),
        # July 2024
        ('2024-07-02', 'Inbound',  'Stainless Bolts M12',     720),
        ('2024-07-05', 'Outbound', 'PVC Tubing 25mm',         195),
        ('2024-07-09', 'Inbound',  'Industrial Bearings 6205',440),
        ('2024-07-12', 'Outbound', 'Stainless Bolts M12',     500),
        ('2024-07-16', 'Inbound',  'Hydraulic Fittings M20',  250),
        ('2024-07-19', 'Outbound', 'Industrial Bearings 6205',310),
        ('2024-07-23', 'Inbound',  'Rubber Gaskets 50mm',     430),
        ('2024-07-27', 'Outbound', 'Hydraulic Fittings M20',  170),
        # August 2024
        ('2024-08-01', 'Inbound',  'Steel Pipes DN50',        340),
        ('2024-08-06', 'Outbound', 'Rubber Gaskets 50mm',     260),
        ('2024-08-09', 'Inbound',  'PVC Tubing 25mm',         380),
        ('2024-08-13', 'Outbound', 'Steel Pipes DN50',        280),
        ('2024-08-16', 'Inbound',  'Stainless Bolts M12',     550),
        ('2024-08-20', 'Outbound', 'PVC Tubing 25mm',         230),
        ('2024-08-23', 'Inbound',  'Industrial Bearings 6205',390),
        ('2024-08-27', 'Outbound', 'Stainless Bolts M12',     410),
        # September 2024
        ('2024-09-03', 'Inbound',  'Hydraulic Fittings M20',  280),
        ('2024-09-06', 'Outbound', 'Industrial Bearings 6205',220),
        ('2024-09-10', 'Inbound',  'Rubber Gaskets 50mm',     470),
        ('2024-09-13', 'Outbound', 'Hydraulic Fittings M20',  200),
        ('2024-09-17', 'Inbound',  'Steel Pipes DN50',        310),
        ('2024-09-20', 'Outbound', 'Rubber Gaskets 50mm',     350),
        ('2024-09-24', 'Inbound',  'PVC Tubing 25mm',         290),
        ('2024-09-27', 'Outbound', 'Steel Pipes DN50',        240),
        # October 2024
        ('2024-10-02', 'Inbound',  'Stainless Bolts M12',     680),
        ('2024-10-05', 'Outbound', 'PVC Tubing 25mm',         210),
        ('2024-10-08', 'Inbound',  'Industrial Bearings 6205',460),
        ('2024-10-11', 'Outbound', 'Stainless Bolts M12',     490),
        ('2024-10-15', 'Inbound',  'Hydraulic Fittings M20',  320),
        ('2024-10-18', 'Outbound', 'Industrial Bearings 6205',290),
        ('2024-10-22', 'Inbound',  'Rubber Gaskets 50mm',     520),
        ('2024-10-25', 'Outbound', 'Hydraulic Fittings M20',  240),
        # November 2024
        ('2024-11-01', 'Inbound',  'Steel Pipes DN50',        380),
        ('2024-11-05', 'Outbound', 'Rubber Gaskets 50mm',     300),
        ('2024-11-08', 'Inbound',  'PVC Tubing 25mm',         420),
        ('2024-11-12', 'Outbound', 'Steel Pipes DN50',        295),
        ('2024-11-15', 'Inbound',  'Stainless Bolts M12',     610),
        ('2024-11-19', 'Outbound', 'PVC Tubing 25mm',         280),
        ('2024-11-22', 'Inbound',  'Industrial Bearings 6205',400),
        ('2024-11-26', 'Outbound', 'Stainless Bolts M12',     430),
        # December 2024
        ('2024-12-03', 'Inbound',  'Hydraulic Fittings M20',  360),
        ('2024-12-06', 'Outbound', 'Industrial Bearings 6205',250),
        ('2024-12-10', 'Inbound',  'Rubber Gaskets 50mm',     540),
        ('2024-12-13', 'Outbound', 'Hydraulic Fittings M20',  290),
        ('2024-12-17', 'Inbound',  'Steel Pipes DN50',        420),
        ('2024-12-20', 'Outbound', 'Rubber Gaskets 50mm',     380),
        ('2024-12-24', 'Inbound',  'PVC Tubing 25mm',         310),
        ('2024-12-27', 'Outbound', 'Steel Pipes DN50',        320),
    ]

    for r, (date_val, mov_type, product, units) in enumerate(movement_data, 2):
        ws1.cell(row=r, column=1, value=date_val)
        ws1.cell(row=r, column=2, value=mov_type)
        ws1.cell(row=r, column=3, value=product)
        ws1.cell(row=r, column=4, value=units)

    # --- Sheet2: Empty (no summary table yet — that is the task) ---
    ws2 = wb.create_sheet('Sheet2')
    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
