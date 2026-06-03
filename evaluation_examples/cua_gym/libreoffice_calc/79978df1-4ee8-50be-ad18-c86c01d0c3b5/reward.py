"""
Reward Script: Monthly warehouse summary table in Sheet2
Task ID: osworld_calc_sheet2_summary_table_003
Domain: libreoffice_calc

Task: Create a table in Sheet2 showing total inbound units, total outbound units,
and net inventory change per month from Sheet1 data. Sheet2 should have:
  - Header row: Month, Inbound Units, Outbound Units, Net Change
  - 12 data rows (Jan 2024 - Dec 2024)
  - Calculations using formulas referencing Sheet1 (SUMIFS/SUMPRODUCT or equivalent)
  - Net Change = Inbound - Outbound

Scoring:
  Component 1: Sheet2 has correct header row (4 columns) — 0.20 pts
  Component 2: Sheet2 has exactly 12 data rows covering all 12 months — 0.30 pts
  Component 3: Inbound/Outbound columns contain formulas/values derived from Sheet1 — 0.25 pts
  Component 4: Net Change column matches Inbound - Outbound for each row — 0.25 pts
Total: 1.0
"""

import os
from datetime import datetime

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_sheet2_summary_table_003'


def compute_expected_from_sheet1(ws1):
    """
    Compute expected monthly inbound/outbound totals from Sheet1.
    Returns dict: month_num -> {'inbound': int, 'outbound': int}
    """
    monthly_inbound = {}
    monthly_outbound = {}

    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=4, values_only=True):
        date, mtype, product, units = row
        if date is None or mtype is None or units is None:
            continue
        if isinstance(date, str):
            try:
                date = datetime.strptime(date, '%Y-%m-%d')
            except Exception:
                continue
        try:
            month_num = date.month
        except Exception:
            continue
        units_val = int(units)
        mtype_str = str(mtype).strip()
        if mtype_str == 'Inbound':
            monthly_inbound[month_num] = monthly_inbound.get(month_num, 0) + units_val
        elif mtype_str == 'Outbound':
            monthly_outbound[month_num] = monthly_outbound.get(month_num, 0) + units_val

    result = {}
    for m in range(1, 13):
        result[m] = {
            'inbound': monthly_inbound.get(m, 0),
            'outbound': monthly_outbound.get(m, 0),
        }
    return result


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print('CRITICAL: Cannot load file %s: %s' % (file_path, e))
        print('REWARD: 0.0')
        return 0.0

    # Precondition: Sheet2 must exist
    if 'Sheet2' not in wb.sheetnames:
        print('FAIL: Sheet2 does not exist in the workbook')
        print('\nScore: 0.0/1.0')
        print('REWARD: 0.0')
        return 0.0

    # Precondition: Sheet1 must exist for reference computation
    if 'Sheet1' not in wb.sheetnames:
        print('FAIL: Sheet1 does not exist; cannot verify calculations')
        print('\nScore: 0.0/1.0')
        print('REWARD: 0.0')
        return 0.0

    ws1 = wb['Sheet1']
    ws2 = wb['Sheet2']

    # Compute expected values from Sheet1
    try:
        expected = compute_expected_from_sheet1(ws1)
    except Exception as e:
        print('ERROR: Could not compute expected values from Sheet1: %s' % e)
        expected = None

    # ----------------------------------------------------------------
    # Component 1: Sheet2 has correct header row (0.20 points)
    # The header row must have columns: Month, Inbound Units, Outbound Units, Net Change
    # This is a task-introduced change (initial Sheet2 is empty)
    # ----------------------------------------------------------------
    try:
        expected_headers = ['Month', 'Inbound Units', 'Outbound Units', 'Net Change']
        row1 = [ws2.cell(row=1, column=c).value for c in range(1, 5)]
        # Require non-None values for all 4 header cells
        if any(v is None for v in row1):
            print('FAIL: Component 1 — Header row has None values: %s' % row1)
            matched_headers = 0
        else:
            # Normalize: strip whitespace and compare case-insensitively
            actual_headers_norm = [str(v).strip().lower() for v in row1]
            expected_headers_norm = [h.lower() for h in expected_headers]

            matched_headers = 0
            for exp, act in zip(expected_headers_norm, actual_headers_norm):
                if exp in act or act in exp:
                    matched_headers += 1

        if matched_headers == 4:
            print('PASS: Component 1 — Header row has all 4 expected columns: %s (0.20 pts)' % row1)
            total_score += 0.20
        elif matched_headers >= 2:
            print('PARTIAL: Component 1 — Header row partially matches (%d/4 columns): %s' % (matched_headers, row1))
            # No partial credit for header; it must be correct
        else:
            print('FAIL: Component 1 — Header row incorrect. Expected %s, found %s' % (expected_headers, row1))
    except Exception as e:
        print('ERROR: Component 1 — %s' % e)

    # ----------------------------------------------------------------
    # Component 2: Sheet2 has exactly 12 data rows covering all 12 months (0.30 points)
    # The initial Sheet2 is empty, so any rows here represent the task change
    # ----------------------------------------------------------------
    try:
        MONTH_PATTERNS = [
            'jan', 'feb', 'mar', 'apr', 'may', 'jun',
            'jul', 'aug', 'sep', 'oct', 'nov', 'dec'
        ]

        data_rows = []
        for r in range(2, ws2.max_row + 1):
            month_cell_val = ws2.cell(row=r, column=1).value
            if month_cell_val is not None:
                data_rows.append((r, str(month_cell_val).strip()))

        print('INFO: Component 2 — Found %d data rows in Sheet2' % len(data_rows))

        # Check that all 12 months are covered
        found_months = set()
        for _, month_str in data_rows:
            month_lower = month_str.lower()
            for i, pat in enumerate(MONTH_PATTERNS):
                if pat in month_lower:
                    found_months.add(i + 1)
                    break

        if len(data_rows) == 12 and len(found_months) == 12:
            print('PASS: Component 2 — 12 data rows present, all 12 months covered (0.30 pts)')
            total_score += 0.30
        elif len(data_rows) == 12:
            print('PARTIAL: Component 2 — 12 rows found but only %d distinct months identified' % len(found_months))
            # Still count as pass if row count is correct (month label format might differ)
            if len(found_months) >= 6:
                print('INFO: Awarding partial credit for 12 rows with %d months recognized' % len(found_months))
                total_score += 0.15
        elif len(data_rows) > 0:
            print('FAIL: Component 2 — Expected 12 data rows, found %d rows. Months: %s' % (len(data_rows), found_months))
        else:
            print('FAIL: Component 2 — No data rows found in Sheet2')
    except Exception as e:
        print('ERROR: Component 2 — %s' % e)

    # ----------------------------------------------------------------
    # Component 3: Inbound/Outbound columns contain formulas or values
    # derived from Sheet1 (not just hardcoded zeros) (0.25 points)
    # ----------------------------------------------------------------
    try:
        formula_count = 0
        nonzero_count = 0
        sheet1_ref_count = 0
        total_cells_checked = 0

        for r in range(2, min(ws2.max_row + 1, 14)):  # rows 2-13
            ib_cell = ws2.cell(row=r, column=2)  # Inbound Units
            ob_cell = ws2.cell(row=r, column=3)  # Outbound Units
            for cell in [ib_cell, ob_cell]:
                total_cells_checked += 1
                val = cell.value
                if val is None:
                    continue
                if isinstance(val, str) and val.startswith('='):
                    formula_count += 1
                    if 'Sheet1' in val or 'sheet1' in val.lower():
                        sheet1_ref_count += 1
                elif isinstance(val, (int, float)) and val != 0:
                    nonzero_count += 1

        print('INFO: Component 3 — formula_count=%d, sheet1_ref_count=%d, nonzero_count=%d (of %d cells)' % (
            formula_count, sheet1_ref_count, nonzero_count, total_cells_checked))

        if sheet1_ref_count >= 12:
            # At least half the Inbound/Outbound cells have Sheet1-referencing formulas
            print('PASS: Component 3 — Inbound/Outbound columns have Sheet1-referencing formulas (%d cells) (0.25 pts)' % sheet1_ref_count)
            total_score += 0.25
        elif formula_count >= 12:
            print('PASS: Component 3 — Inbound/Outbound columns have formulas (%d cells) (0.25 pts)' % formula_count)
            total_score += 0.25
        elif nonzero_count >= 12:
            # Hardcoded non-zero values also acceptable if they match Sheet1 calculations
            print('INFO: Component 3 — Values found (not formulas), checking if they match Sheet1 computation')
            if expected is not None:
                match_count = 0
                for r in range(2, min(ws2.max_row + 1, 14)):
                    month_num = r - 1  # row 2=month 1, etc.
                    ib_val = ws2.cell(row=r, column=2).value
                    ob_val = ws2.cell(row=r, column=3).value
                    exp_ib = expected.get(month_num, {}).get('inbound', 0)
                    exp_ob = expected.get(month_num, {}).get('outbound', 0)
                    if ib_val == exp_ib and ob_val == exp_ob:
                        match_count += 1
                if match_count >= 6:
                    print('PASS: Component 3 — Values match Sheet1 computations (%d/12 rows match) (0.25 pts)' % match_count)
                    total_score += 0.25
                else:
                    print('FAIL: Component 3 — Only %d/12 rows match expected Sheet1 values' % match_count)
            else:
                print('FAIL: Component 3 — Non-zero values found but cannot verify against Sheet1')
        else:
            print('FAIL: Component 3 — Inbound/Outbound columns appear to have no meaningful content (%d formulas, %d non-zero values)' % (formula_count, nonzero_count))
    except Exception as e:
        print('ERROR: Component 3 — %s' % e)

    # ----------------------------------------------------------------
    # Component 4: Net Change column correctly reflects Inbound - Outbound (0.25 points)
    # Verified by computing from Sheet1 data directly
    # ----------------------------------------------------------------
    try:
        net_change_correct = 0
        net_change_total = 0

        if expected is not None:
            for r in range(2, min(ws2.max_row + 1, 14)):
                month_num = r - 1  # row 2 = month 1 (Jan), etc.
                net_cell = ws2.cell(row=r, column=4)
                net_val = net_cell.value

                exp_ib = expected[month_num]['inbound']
                exp_ob = expected[month_num]['outbound']
                exp_net = exp_ib - exp_ob

                if net_val is None:
                    net_change_total += 1
                    print('FAIL row %d: Net Change is None, expected %d' % (r, exp_net))
                    continue

                net_change_total += 1

                # If it's a formula like =B2-C2, we verify the formula structure
                if isinstance(net_val, str) and net_val.startswith('='):
                    # Check it's a subtraction formula referencing same row's B-C
                    import re
                    col_b = 'B%d' % r
                    col_c = 'C%d' % r
                    formula_clean = net_val.replace(' ', '').upper()
                    if col_b.upper() in formula_clean and col_c.upper() in formula_clean and '-' in formula_clean:
                        net_change_correct += 1
                    else:
                        print('FAIL row %d: Net Change formula "%s" does not look like Inbound-Outbound' % (r, net_val))
                elif isinstance(net_val, (int, float)):
                    # Numeric value — check it matches expected
                    if abs(float(net_val) - exp_net) < 0.5:
                        net_change_correct += 1
                    else:
                        print('FAIL row %d: Net Change value %s, expected %d' % (r, net_val, exp_net))
                else:
                    print('FAIL row %d: Net Change unexpected type: %s (%s)' % (r, type(net_val).__name__, net_val))

            print('INFO: Component 4 — Net Change correct for %d/%d rows' % (net_change_correct, net_change_total))

            if net_change_total > 0 and net_change_correct == net_change_total:
                print('PASS: Component 4 — All Net Change values/formulas are correct (0.25 pts)')
                total_score += 0.25
            elif net_change_total > 0 and net_change_correct >= 9:
                # At least 9/12 rows correct — award proportional credit
                partial = round(0.25 * net_change_correct / net_change_total, 4)
                print('PARTIAL: Component 4 — %d/%d Net Change rows correct' % (net_change_correct, net_change_total))
                if partial > 0:
                    total_score += partial
            else:
                print('FAIL: Component 4 — Only %d/%d Net Change rows correct' % (net_change_correct, net_change_total))
        else:
            # No expected data available — try formula-only check
            formula_net_count = 0
            for r in range(2, min(ws2.max_row + 1, 14)):
                net_val = ws2.cell(row=r, column=4).value
                if isinstance(net_val, str) and net_val.startswith('=') and '-' in net_val:
                    formula_net_count += 1
            if formula_net_count >= 12:
                print('PASS: Component 4 — Net Change column has subtraction formulas (%d rows) (0.25 pts)' % formula_net_count)
                total_score += 0.25
            else:
                print('FAIL: Component 4 — Net Change column lacks proper subtraction formulas')
    except Exception as e:
        print('ERROR: Component 4 — %s' % e)

    final_score = min(total_score, 1.0)
    print('\nScore: %.4f/1.0' % total_score)
    print('REWARD: %.1f' % final_score)
    return final_score


# Entry point: test against canonical artifact path
file_path = '%s/%s.xlsx' % (WORKDIR, TASK_ID)
if not os.path.exists(file_path):
    print('File not found: %s' % file_path)
    print('REWARD: 0.0')
else:
    verify_task(file_path)
