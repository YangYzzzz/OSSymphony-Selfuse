"""
Initial Setup: Paste Special Transpose - horizontal months to vertical column
Task ID: calc_gsi_040
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_040'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Revenue Tracking"

    # Row 1: Month labels arranged horizontally (A1:L1)
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    for col, month in enumerate(months, 1):
        cell = ws.cell(row=1, column=col, value=month)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        cell.font = Font(bold=True, size=11, color="FFFFFF")

    # Row 2: Revenue figures for North region
    ws.cell(row=2, column=1, value="North Region Revenue").font = Font(italic=True)
    # Merge the label across a few cells for clarity - actually let's keep it simple
    # Row 2 header label in a merged area would be complex; use a separate approach
    # Actually, let's use a cleaner layout:

    # Clear row 2 and redo: Put region labels and data starting row 3
    ws.cell(row=2, column=1, value=None)

    # Row 2: North Region revenue by month
    north_revenue = [45230, 52100, 48750, 61200, 55800, 67300,
                     72150, 69400, 58900, 63700, 71200, 84500]
    for col, val in enumerate(north_revenue, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.number_format = '#,##0'

    # Row 3: South Region revenue by month
    south_revenue = [38100, 41500, 44200, 39800, 47600, 51200,
                     48900, 53100, 46700, 50200, 55800, 62300]
    for col, val in enumerate(south_revenue, 1):
        cell = ws.cell(row=3, column=col, value=val)
        cell.number_format = '#,##0'

    # Row 4: East Region revenue by month
    east_revenue = [52400, 49800, 55100, 58700, 61200, 54300,
                    57800, 63400, 59100, 66200, 70500, 75800]
    for col, val in enumerate(east_revenue, 1):
        cell = ws.cell(row=4, column=col, value=val)
        cell.number_format = '#,##0'

    # Row 5: West Region revenue by month
    west_revenue = [41700, 43200, 46800, 50100, 48300, 52700,
                    55400, 51900, 49600, 54800, 58100, 63200]
    for col, val in enumerate(west_revenue, 1):
        cell = ws.cell(row=5, column=col, value=val)
        cell.number_format = '#,##0'

    # Row 6: Total
    for col in range(1, 13):
        cell = ws.cell(row=6, column=col)
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(col)
        cell.value = f'=SUM({col_letter}2:{col_letter}5)'
        cell.font = Font(bold=True)
        cell.number_format = '#,##0'

    # Set column widths
    for col in range(1, 13):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = 12

    # Add a second sheet with summary info to add complexity
    ws2 = wb.create_sheet("Regions")
    ws2.cell(row=1, column=1, value="Region").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Manager").font = Font(bold=True)
    ws2.cell(row=1, column=3, value="Target").font = Font(bold=True)
    ws2.cell(row=1, column=4, value="Status").font = Font(bold=True)

    regions_data = [
        ["North", "Sarah Chen", 700000, "On Track"],
        ["South", "Marcus Johnson", 580000, "Behind"],
        ["East", "Priya Patel", 720000, "Ahead"],
        ["West", "David Kim", 600000, "On Track"],
    ]
    for r, row_data in enumerate(regions_data, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 14
    ws2.column_dimensions['D'].width = 12

    # NOTE: Row 1 has months horizontally. The task asks to transpose them
    # to a vertical column. The initial file must NOT have a vertical month column.

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
