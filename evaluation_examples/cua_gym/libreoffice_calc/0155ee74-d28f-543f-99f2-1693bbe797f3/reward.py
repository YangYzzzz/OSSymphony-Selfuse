"""
Reward Script: Transpose horizontal month labels to vertical column using Paste Special
Task ID: calc_gsi_040
Domain: libreoffice_calc
Scoring:
  Precondition gate: Original horizontal row 1 data (A1:L1) must be intact
  Component 1 (0.6): All 12 month labels present vertically in column A starting at row 8
  Component 2 (0.4): Month labels are in correct order (Jan..Dec) in A8:A19
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_040'

EXPECTED_MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                   'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Revenue Tracking']

    # Precondition gate: Original horizontal row 1 data must be intact
    # This is NOT scored -- it's a prerequisite. If original data is corrupted, score 0.
    try:
        horizontal_values = []
        for c in range(1, 13):
            val = ws.cell(row=1, column=c).value
            if val is not None:
                horizontal_values.append(str(val).strip())
            else:
                horizontal_values.append(None)

        if horizontal_values != EXPECTED_MONTHS:
            print(f"PRECONDITION FAIL: Original row 1 data corrupted. Found: {horizontal_values}")
            print("REWARD: 0.0")
            return 0.0
        else:
            print(f"PRECONDITION OK: Original horizontal months A1:L1 intact")
    except Exception as e:
        print(f"ERROR: Precondition check — {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: All 12 month labels present vertically in column A rows 8-19 (0.6 points)
    # This checks that the transpose actually happened -- these cells are empty in the initial file
    try:
        vertical_months = []
        for r in range(8, 20):
            val = ws.cell(row=r, column=1).value
            if val is not None:
                vertical_months.append(str(val).strip())

        match_count = 0
        for month in EXPECTED_MONTHS:
            if month in vertical_months:
                match_count += 1

        if match_count == 12:
            print(f"PASS: Component 1 — All 12 months found vertically in A8:A19 (0.6 pts)")
            total_score += 0.6
        elif match_count >= 1:
            partial = round(0.6 * (match_count / 12), 2)
            print(f"PARTIAL: Component 1 — {match_count}/12 months found vertically ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No months found vertically. Found: {vertical_months}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Month labels are in correct order Jan..Dec in A8:A19 (0.4 points)
    # Verifies the transpose preserved the exact sequence order
    try:
        vertical_values = []
        for r in range(8, 20):
            val = ws.cell(row=r, column=1).value
            if val is not None:
                vertical_values.append(str(val).strip())
            else:
                vertical_values.append(None)

        if len(vertical_values) == 12 and vertical_values == EXPECTED_MONTHS:
            print(f"PASS: Component 2 — Months in correct order Jan..Dec in A8:A19 (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 2 — Expected {EXPECTED_MONTHS}, found {vertical_values}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
