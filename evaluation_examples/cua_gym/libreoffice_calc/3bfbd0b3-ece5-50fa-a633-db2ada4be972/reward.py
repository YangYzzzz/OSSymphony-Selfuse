"""
Reward Script: Copy 'Invoice Template' sheet and rename to 'Invoice #1042', placed at end.
Task ID: calc_ps_061
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.25): 'Invoice #1042' sheet exists
  - Component 2 (0.15): Sheet is at the last position in the workbook
  - Component 3 (0.15): Sheet dimensions match 'Invoice Template'
  - Component 4 (0.25): All cell values match the template
  - Component 5 (0.20): Formulas are preserved from the template
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_061'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice document."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    sheet_names = wb.sheetnames

    # Component 1: 'Invoice #1042' sheet exists (0.25 points)
    try:
        if 'Invoice #1042' in sheet_names:
            print(f"PASS: Component 1 — 'Invoice #1042' sheet exists (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — 'Invoice #1042' not found. Sheets: {sheet_names}")
            # If the sheet doesn't exist, no further checks are meaningful
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: 'Invoice #1042' is the LAST sheet in the workbook (0.15 points)
    try:
        idx = sheet_names.index('Invoice #1042')
        if idx == len(sheet_names) - 1:
            print(f"PASS: Component 2 — 'Invoice #1042' is the last sheet (index {idx} of {len(sheet_names)}) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 — 'Invoice #1042' at index {idx}, expected last (index {len(sheet_names) - 1})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Precondition: 'Invoice Template' must exist for comparison
    if 'Invoice Template' not in sheet_names:
        print(f"PRECONDITION FAIL: 'Invoice Template' sheet missing. Cannot compare.")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    template = wb['Invoice Template']
    copy_sheet = wb['Invoice #1042']

    # Component 3: Sheet dimensions match template (0.15 points)
    try:
        t_rows = template.max_row
        t_cols = template.max_column
        c_rows = copy_sheet.max_row
        c_cols = copy_sheet.max_column
        if t_rows == c_rows and t_cols == c_cols:
            print(f"PASS: Component 3 — Dimensions match: {c_rows}x{c_cols} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 — Template {t_rows}x{t_cols}, Copy {c_rows}x{c_cols}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: All cell values match the template (0.25 points)
    try:
        mismatches = 0
        total_cells = 0
        for row in range(1, template.max_row + 1):
            for col in range(1, template.max_column + 1):
                t_val = template.cell(row=row, column=col).value
                c_val = copy_sheet.cell(row=row, column=col).value
                if t_val is not None or c_val is not None:
                    total_cells += 1
                    if t_val != c_val:
                        mismatches += 1

        if total_cells > 0 and mismatches == 0:
            print(f"PASS: Component 4 — All {total_cells} non-empty cells match (0.25 pts)")
            total_score += 0.25
        elif total_cells > 0:
            # Partial credit: proportional to matching cells
            match_ratio = (total_cells - mismatches) / total_cells
            partial = round(0.25 * match_ratio, 2)
            print(f"FAIL: Component 4 — {mismatches}/{total_cells} cells differ (partial: {partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — No non-empty cells found in either sheet")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Formulas preserved from template (0.20 points)
    try:
        formula_count = 0
        formula_match = 0
        for row in range(1, template.max_row + 1):
            for col in range(1, template.max_column + 1):
                t_val = template.cell(row=row, column=col).value
                if isinstance(t_val, str) and t_val.startswith('='):
                    formula_count += 1
                    c_val = copy_sheet.cell(row=row, column=col).value
                    if isinstance(c_val, str) and c_val.upper().replace(" ", "") == t_val.upper().replace(" ", ""):
                        formula_match += 1

        if formula_count > 0 and formula_match == formula_count:
            print(f"PASS: Component 5 — All {formula_count} formulas preserved (0.20 pts)")
            total_score += 0.20
        elif formula_count > 0:
            ratio = formula_match / formula_count
            partial = round(0.20 * ratio, 2)
            print(f"FAIL: Component 5 — {formula_match}/{formula_count} formulas match (partial: {partial} pts)")
            total_score += partial
        else:
            # No formulas in template — this component is N/A, award full points
            print(f"PASS: Component 5 — No formulas in template, N/A (0.20 pts)")
            total_score += 0.20
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
