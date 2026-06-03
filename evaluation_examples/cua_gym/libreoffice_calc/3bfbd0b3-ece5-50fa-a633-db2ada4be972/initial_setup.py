"""
Initial Setup: Copy the 'Invoice Template' sheet and rename the copy to 'Invoice #1042'
Task ID: calc_ps_061
Domain: libreoffice_calc

Initial state: Workbook with 'Invoice Template' (formatted invoice layout with formulas)
and 'Clients' sheet. No 'Invoice #1042' sheet yet.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_061'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def build_invoice_template(ws):
    """Build the Invoice Template sheet with formatted layout and formulas."""
    # --- Styles ---
    header_font = Font(name="Arial", size=18, bold=True, color="1F4E79")
    subheader_font = Font(name="Arial", size=11, bold=True, color="2E75B6")
    label_font = Font(name="Arial", size=10, bold=True)
    data_font = Font(name="Arial", size=10)
    table_header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    table_header_fill = PatternFill(start_color="FF2E75B6", end_color="FF2E75B6", fill_type="solid")
    light_fill = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")
    currency_fmt = '$#,##0.00'
    thin_border = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999"),
    )

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16

    # --- Company Header (rows 1-3) ---
    ws.merge_cells("A1:E1")
    ws["A1"] = "Pinnacle Solutions Inc."
    ws["A1"].font = header_font
    ws["A1"].alignment = Alignment(horizontal="left")

    ws.merge_cells("A2:E2")
    ws["A2"] = "1250 Technology Drive, Suite 400, San Francisco, CA 94107"
    ws["A2"].font = Font(name="Arial", size=9, color="555555")

    ws.merge_cells("A3:E3")
    ws["A3"] = "Phone: (415) 555-0198  |  Email: billing@pinnaclesolutions.com"
    ws["A3"].font = Font(name="Arial", size=9, color="555555")

    # --- Invoice Title (row 5) ---
    ws.merge_cells("A5:E5")
    ws["A5"] = "INVOICE"
    ws["A5"].font = Font(name="Arial", size=14, bold=True, color="2E75B6")
    ws["A5"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[5].height = 28

    # --- Invoice Details (rows 7-10) ---
    details = [
        (7, "Invoice Number:", "[Auto-Generated]"),
        (8, "Invoice Date:", ""),
        (9, "Due Date:", ""),
        (10, "Payment Terms:", "Net 30"),
    ]
    for row_num, label, value in details:
        ws.cell(row=row_num, column=1, value=label).font = label_font
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
        ws.cell(row=row_num, column=3, value=value).font = data_font
        ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=5)

    # --- Bill To (rows 12-15) ---
    ws.merge_cells("A12:B12")
    ws["A12"] = "Bill To:"
    ws["A12"].font = subheader_font
    bill_to = [
        (13, "[Client Name]"),
        (14, "[Client Address]"),
        (15, "[City, State ZIP]"),
    ]
    for row_num, value in bill_to:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
        ws.cell(row=row_num, column=1, value=value).font = data_font

    # --- Line Items Table Header (row 17) ---
    table_headers = ["#", "Description", "Qty", "Unit Price", "Amount"]
    for col, header in enumerate(table_headers, 1):
        cell = ws.cell(row=17, column=col, value=header)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    ws.row_dimensions[17].height = 22

    # --- Line Items (rows 18-25, 8 template rows) ---
    for r in range(18, 26):
        ws.cell(row=r, column=1, value=r - 17).font = data_font
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2).font = data_font
        ws.cell(row=r, column=3).font = data_font
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=4).font = data_font
        ws.cell(row=r, column=4).number_format = currency_fmt
        # Amount formula: Qty * Unit Price
        ws.cell(row=r, column=5, value=f'=IF(AND(C{r}<>"",D{r}<>""),C{r}*D{r},"")')
        ws.cell(row=r, column=5).font = data_font
        ws.cell(row=r, column=5).number_format = currency_fmt
        # Alternate row shading
        if (r - 18) % 2 == 0:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = light_fill
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = thin_border

    # Sample data for a few rows to make it non-trivial
    sample_items = [
        ("Website Redesign - Phase 1", 1, 4500.00),
        ("SEO Optimization Package", 1, 1200.00),
        ("Content Writing (per article)", 5, 150.00),
        ("Social Media Management (monthly)", 2, 800.00),
        ("Email Campaign Setup", 1, 650.00),
    ]
    for i, (desc, qty, price) in enumerate(sample_items):
        row = 18 + i
        ws.cell(row=row, column=2, value=desc)
        ws.cell(row=row, column=3, value=qty)
        ws.cell(row=row, column=4, value=price)

    # --- Totals Section (rows 27-29) ---
    ws.merge_cells("C27:D27")
    ws["C27"] = "Subtotal:"
    ws["C27"].font = Font(name="Arial", size=10, bold=True)
    ws["C27"].alignment = Alignment(horizontal="right")
    ws["E27"] = '=SUMPRODUCT((E18:E25<>"")*E18:E25)'
    ws["E27"].font = Font(name="Arial", size=10, bold=True)
    ws["E27"].number_format = currency_fmt
    ws["E27"].border = thin_border

    ws.merge_cells("C28:D28")
    ws["C28"] = "Tax (8.5%):"
    ws["C28"].font = label_font
    ws["C28"].alignment = Alignment(horizontal="right")
    ws["E28"] = '=E27*0.085'
    ws["E28"].font = data_font
    ws["E28"].number_format = currency_fmt
    ws["E28"].border = thin_border

    ws.merge_cells("C29:D29")
    ws["C29"] = "Total Due:"
    ws["C29"].font = Font(name="Arial", size=12, bold=True, color="1F4E79")
    ws["C29"].alignment = Alignment(horizontal="right")
    ws["E29"] = '=E27+E28'
    ws["E29"].font = Font(name="Arial", size=12, bold=True, color="1F4E79")
    ws["E29"].number_format = currency_fmt
    total_border = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="double", color="2E75B6"),
        bottom=Side(style="double", color="2E75B6"),
    )
    ws["E29"].border = total_border

    # --- Notes (row 31) ---
    ws.merge_cells("A31:E31")
    ws["A31"] = "Notes:"
    ws["A31"].font = subheader_font

    ws.merge_cells("A32:E33")
    ws["A32"] = "Thank you for your business! Please make payment within the specified terms. For questions regarding this invoice, contact billing@pinnaclesolutions.com."
    ws["A32"].font = Font(name="Arial", size=9, color="666666")
    ws["A32"].alignment = Alignment(wrap_text=True, vertical="top")


def build_clients_sheet(ws):
    """Build the Clients sheet with client data."""
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 18

    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2E75B6", end_color="FF2E75B6", fill_type="solid")
    data_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    headers = ["ID", "Company Name", "Contact Person", "Email", "Phone", "Account Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    clients = [
        (1001, "TechVentures Corp", "Laura Mitchell", "l.mitchell@techventures.com", "(212) 555-0134", "Active"),
        (1002, "GreenLeaf Organics", "David Park", "david.park@greenleaf.org", "(503) 555-0267", "Active"),
        (1003, "Atlas Manufacturing", "Rebecca Torres", "r.torres@atlasmfg.com", "(312) 555-0189", "Active"),
        (1004, "Skyline Digital Media", "James Whitfield", "j.whitfield@skylinedm.com", "(415) 555-0342", "Inactive"),
        (1005, "Meridian Health Group", "Priya Sharma", "p.sharma@meridianhealth.com", "(617) 555-0456", "Active"),
        (1006, "Coastal Engineering LLC", "Nathan Brooks", "n.brooks@coastaleng.com", "(858) 555-0523", "Active"),
        (1007, "Summit Financial Advisors", "Angela Kim", "a.kim@summitfa.com", "(202) 555-0678", "Active"),
        (1008, "Redwood Logistics", "Carlos Mendez", "c.mendez@redwoodlog.com", "(713) 555-0791", "Pending"),
        (1009, "BrightPath Education", "Emily Watson", "e.watson@brightpath.edu", "(404) 555-0845", "Active"),
        (1010, "Pacific Trade Solutions", "Ryan O'Connor", "r.oconnor@pacifictrade.com", "(206) 555-0912", "Active"),
        (1011, "Nova Creative Agency", "Sofia Alvarez", "s.alvarez@novacreative.com", "(310) 555-0167", "Active"),
        (1012, "Irongate Construction", "Thomas Henderson", "t.henderson@irongate.com", "(469) 555-0234", "Inactive"),
    ]

    for r, row_data in enumerate(clients, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = thin_border
            if c == 1:
                cell.alignment = Alignment(horizontal="center")


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Invoice Template ---
    ws1 = wb.active
    ws1.title = "Invoice Template"
    build_invoice_template(ws1)

    # --- Sheet 2: Clients ---
    ws2 = wb.create_sheet("Clients")
    build_clients_sheet(ws2)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
