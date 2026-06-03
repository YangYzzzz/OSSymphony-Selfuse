"""
Initial Setup: Fix wildcard VLOOKUP formula that returns #N/A
Task ID: calc_tbl_087
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_087'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Employee Data ---
    ws1 = wb.active
    ws1.title = 'Employee Data'

    headers = ['Full Name', 'Department', 'Annual Salary']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.font = Font(bold=True, size=11, color="FFFFFF")

    data = [
        ['Maria Garcia', 'Marketing', 68500],
        ['John Smith', 'Engineering', 92000],
        ['Priya Patel', 'Finance', 78000],
        ['David Lee', 'Operations', 71000],
        ['Jane Smith-Jones', 'Human Resources', 85000],
        ['Robert Chen', 'Engineering', 95000],
        ['Emily Watson', 'Marketing', 72500],
        ['Michael Brown', 'Finance', 81000],
        ['Sarah Thompson', 'Operations', 69000],
        ['James Wilson', 'Engineering', 88500],
        ['Laura Smith', 'Finance', 76000],
        ['Daniel Kim', 'Marketing', 73500],
        ['Angela Rivera', 'Human Resources', 82000],
        ['Thomas Anderson', 'Operations', 67500],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            if c == 3:
                cell.number_format = '$#,##0'

    # Set column widths
    ws1.column_dimensions['A'].width = 22
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 16

    # --- Sheet 2: Lookup ---
    ws2 = wb.create_sheet('Lookup')

    ws2.cell(row=1, column=1, value='Search Pattern').font = Font(bold=True)
    ws2.cell(row=1, column=2, value='Department Result').font = Font(bold=True)
    ws2.cell(row=1, column=3, value='Notes').font = Font(bold=True)

    ws2.cell(row=2, column=1, value='*Smith*')

    # BROKEN formula: Using VLOOKUP with wildcards but with match_type=1 (approximate/sorted)
    # instead of match_type=0 (exact match, which is required for wildcards to work).
    # This will return #N/A or wrong results because wildcards only work with exact match mode.
    ws2.cell(row=2, column=2, value='=VLOOKUP("*Smith*",\'Employee Data\'.A:C,2,1)')

    ws2.cell(row=2, column=3, value='Formula returns #N/A - wildcard not working')

    ws2.cell(row=4, column=1, value='Expected:').font = Font(italic=True)
    ws2.cell(row=4, column=2, value='Should return the department of the first name containing "Smith"').font = Font(italic=True)

    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 22
    ws2.column_dimensions['C'].width = 40

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
