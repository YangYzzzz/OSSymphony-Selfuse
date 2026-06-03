"""
Reward Script: Fix VLOOKUP wildcard formula
Task ID: calc_tbl_087
Domain: libreoffice_calc
Scoring:
  Component 1 (0.6 pts): VLOOKUP in B2 uses exact match mode (0 or FALSE)
  Component 2 (0.4 pts): VLOOKUP formula is well-formed with wildcard + exact match
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_087'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def parse_vlookup_args(formula):
    """
    Parse a VLOOKUP formula string into its arguments.
    Returns list of argument strings, or None if not a VLOOKUP.
    Handles nested quotes and sheet references with single quotes.
    """
    if not formula or not isinstance(formula, str):
        return None
    formula_upper = formula.upper().replace(" ", "")
    if not formula_upper.startswith("=VLOOKUP("):
        return None

    # Extract content inside the outer VLOOKUP(...)
    # Find the matching closing paren
    inner = formula[len("=VLOOKUP("):]
    if inner.endswith(")"):
        inner = inner[:-1]
    else:
        return None

    # Split by commas, respecting parentheses and quotes
    args = []
    depth = 0
    in_quote = False
    current = ""
    for ch in inner:
        if ch == '"' and not in_quote:
            in_quote = True
            current += ch
        elif ch == '"' and in_quote:
            in_quote = False
            current += ch
        elif ch == '(' and not in_quote:
            depth += 1
            current += ch
        elif ch == ')' and not in_quote:
            depth -= 1
            current += ch
        elif ch == ',' and depth == 0 and not in_quote:
            args.append(current.strip())
            current = ""
        else:
            current += ch
    if current.strip():
        args.append(current.strip())

    return args


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Lookup sheet must exist
    if 'Lookup' not in wb.sheetnames:
        print("CRITICAL: 'Lookup' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Lookup']
    formula_raw = ws['B2'].value

    print(f"INFO: B2 raw value = {repr(formula_raw)}")

    # Parse the VLOOKUP formula
    args = parse_vlookup_args(str(formula_raw) if formula_raw else "")

    if args is None or len(args) < 4:
        print(f"FAIL: B2 does not contain a valid VLOOKUP with 4 arguments. Value: {repr(formula_raw)}")
        print(f"REWARD: {total_score}")
        return total_score

    print(f"INFO: VLOOKUP args parsed: {args}")

    lookup_value = args[0]
    table_array = args[1]
    col_index = args[2]
    match_type = args[3]

    # Component 1: VLOOKUP uses exact match mode (0.6 points)
    # The critical fix: match_type must be 0 or FALSE (exact match)
    # Wildcards in VLOOKUP only work with exact match mode.
    # Initial file has match_type=1 (approximate), which breaks wildcards.
    try:
        match_type_clean = match_type.strip().upper()
        is_exact_match = match_type_clean in ("0", "FALSE")
        if is_exact_match:
            print(f"PASS: Component 1 -- VLOOKUP match_type is '{match_type}' (exact match) (0.6 pts)")
            total_score += 0.6
        else:
            print(f"FAIL: Component 1 -- VLOOKUP match_type is '{match_type}', expected 0 or FALSE for wildcard support")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Formula is well-formed with wildcard pattern + exact match (0.4 points)
    # Check that the formula uses a wildcard pattern containing "Smith" AND uses exact match
    # This ensures the complete fix: wildcard syntax + correct match mode working together
    try:
        has_wildcard = '"*' in lookup_value and 'Smith' in lookup_value and '*"' in lookup_value
        # Also accept patterns like "*smith*" case-insensitive
        if not has_wildcard:
            has_wildcard = '"*' in lookup_value.lower() and 'smith' in lookup_value.lower()

        has_correct_col = col_index.strip() == "2"
        has_table_ref = "Employee Data" in table_array or "employee data" in table_array.lower()

        if is_exact_match and has_wildcard and has_correct_col and has_table_ref:
            print(f"PASS: Component 2 -- Formula is well-formed: wildcard='{lookup_value}', col_index={col_index}, table_ref has 'Employee Data', exact match (0.4 pts)")
            total_score += 0.4
        else:
            reasons = []
            if not is_exact_match:
                reasons.append(f"match_type '{match_type}' is not exact")
            if not has_wildcard:
                reasons.append(f"lookup_value '{lookup_value}' missing wildcard *Smith* pattern")
            if not has_correct_col:
                reasons.append(f"col_index '{col_index}' should be 2")
            if not has_table_ref:
                reasons.append(f"table_array '{table_array}' missing 'Employee Data' reference")
            print(f"FAIL: Component 2 -- {'; '.join(reasons)}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist GUI state before verification
persist_app_state("libreoffice_calc")

# Run verification
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
