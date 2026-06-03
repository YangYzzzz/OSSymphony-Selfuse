"""
Initial Setup: Price-Volume-Mix Analysis - PVM spreadsheet with raw data
Task ID: calc_fin_price_volume_mix_071
Domain: libreoffice_calc

Creates a spreadsheet with product sales data for PY and CY comparison.
Columns F-J are left empty for the agent to fill in with PVM analysis formulas.
Row 1 headers and data are in columns A-E only.
"""

import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_price_volume_mix_071'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: PVM ---
    ws = wb.active
    ws.title = 'PVM'

    # Row 1 headers (A-E only; F-J intentionally left empty)
    headers = ['Product', 'PY Volume', 'CY Volume', 'PY Price', 'CY Price']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        # NOT bold — the task requires making row 1 bold

    # 7 products with realistic business data
    # Products: consumer electronics segment
    data = [
        # Product,          PY Vol, CY Vol, PY Price,  CY Price
        ['Laptop Pro 15',   4200,   3980,   1249.99,   1299.99],
        ['Wireless Mouse',  18500,  21200,    29.99,     34.99],
        ['USB-C Hub 7-in-1', 9800,  11400,    59.99,     54.99],
        ['Mechanical Keyboard', 6300, 5850,  149.99,    159.99],
        ['4K Monitor 27"',  3100,   3650,    549.99,    579.99],
        ['Noise-Cancel Headphones', 7200, 8900, 199.99, 229.99],
        ['Webcam HD 1080p', 12400,  10800,    79.99,     74.99],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: PVM')
    print('Columns A-E filled with product data for 7 products')
    print('Columns F-J intentionally empty for agent to complete')


create_initial()
