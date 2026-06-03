"""
Reward Script: Price-Volume-Mix Analysis
Task ID: calc_fin_price_volume_mix_071
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.20): New column headers F1:J1 (PY Revenue, CY Revenue, Volume Effect, Price Effect, Total Change)
  - Component 2 (0.25): PVM formulas in F2:J8 (revenue, volume effect, price effect, total change)
  - Component 3 (0.20): Row 9 totals with SUM formulas (F9:J9) and bold formatting on A9, F9:J9
  - Component 4 (0.15): Currency number format on F2:J8 columns
  - Component 5 (0.10): Conditional formatting on H2:H8 and I2:I8 (positive=green, negative=red)
  - Component 6 (0.10): Bar chart present with 2 series referencing H and I columns
"""

import os
import openpyxl
from openpyxl.utils import get_column_letter

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_price_volume_mix_071'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify PVM sheet exists
    if 'PVM' not in wb.sheetnames:
        print("CRITICAL: 'PVM' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['PVM']

    # Component 1: New column headers F1:J1 (0.20 points)
    # These headers should NOT exist in the initial file (cols F-J are empty initially)
    try:
        expected_headers = {
            6: 'PY Revenue',
            7: 'CY Revenue',
            8: 'Volume Effect',
            9: 'Price Effect',
            10: 'Total Change'
        }
        headers_correct = 0
        for col, expected in expected_headers.items():
            actual = ws.cell(row=1, column=col).value
            if actual and str(actual).strip() == expected:
                headers_correct += 1
            else:
                print(f"FAIL: Header {get_column_letter(col)}1 — expected '{expected}', found: {repr(actual)}")

        if headers_correct == 5:
            print(f"PASS: Component 1 — All 5 headers (F1:J1) correct (0.20 pts)")
            total_score += 0.20
        elif headers_correct >= 3:
            partial = 0.10
            print(f"PARTIAL: Component 1 — {headers_correct}/5 headers correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {headers_correct}/5 headers correct (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: PVM formulas in F2:J8 (0.25 points)
    # F2:F8 = =B*D, G2:G8 = =C*E, H2:H8 = =(C-B)*D, I2:I8 = =(E-D)*C, J2:J8 = =G-F
    # Verify formulas exist for product rows (rows 2-8)
    try:
        formula_checks = {
            'F': {'pattern': ['B', 'D'], 'desc': 'PY Revenue (=B*D)'},
            'G': {'pattern': ['C', 'E'], 'desc': 'CY Revenue (=C*E)'},
            'H': {'pattern': ['C', 'B', 'D'], 'desc': 'Volume Effect (=(C-B)*D)'},
            'I': {'pattern': ['E', 'D', 'C'], 'desc': 'Price Effect (=(E-D)*C)'},
            'J': {'pattern': ['G', 'F'], 'desc': 'Total Change (=G-F)'},
        }
        col_map = {'F': 6, 'G': 7, 'H': 8, 'I': 9, 'J': 10}

        formulas_present = 0
        total_formula_checks = 0

        for col_letter, info in formula_checks.items():
            col_num = col_map[col_letter]
            col_ok = 0
            for row in range(2, 9):  # rows 2-8 (7 products)
                total_formula_checks += 1
                cell_val = ws.cell(row=row, column=col_num).value
                if cell_val and isinstance(cell_val, str) and cell_val.startswith('='):
                    formulas_present += 1
                    col_ok += 1
                else:
                    print(f"FAIL: Component 2 — {col_letter}{row} expected formula, found: {repr(cell_val)}")
            if col_ok == 7:
                print(f"PASS: Component 2 — Column {col_letter} all 7 formulas present ({info['desc']})")

        formula_ratio = formulas_present / total_formula_checks if total_formula_checks > 0 else 0
        if formula_ratio >= 0.95:
            print(f"PASS: Component 2 — {formulas_present}/{total_formula_checks} formulas present (0.25 pts)")
            total_score += 0.25
        elif formula_ratio >= 0.70:
            partial = 0.12
            print(f"PARTIAL: Component 2 — {formulas_present}/{total_formula_checks} formulas present ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {formulas_present}/{total_formula_checks} formulas present (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Row 9 totals with SUM formulas (F9:J9) and bold (0.20 points)
    # Row 9 should have SUM formulas in F9:J9, A9 should say 'Total', and relevant cells bold
    try:
        sum_cols = [6, 7, 8, 9, 10]  # F, G, H, I, J
        sum_formulas_ok = 0
        for col_num in sum_cols:
            cell_val = ws.cell(row=9, column=col_num).value
            if cell_val and isinstance(cell_val, str) and 'SUM' in cell_val.upper():
                sum_formulas_ok += 1
            else:
                print(f"FAIL: Component 3 — {get_column_letter(col_num)}9 expected SUM formula, found: {repr(cell_val)}")

        # Check bold on A9 and F9:J9
        bold_count = 0
        bold_cells = [1, 6, 7, 8, 9, 10]  # A9, F9:J9
        for col_num in bold_cells:
            if ws.cell(row=9, column=col_num).font.bold:
                bold_count += 1

        sums_ok = sum_formulas_ok == 5
        bold_ok = bold_count >= 4  # most bold cells present

        if sums_ok and bold_ok:
            print(f"PASS: Component 3 — Row 9 has all SUM formulas and bold formatting (0.20 pts)")
            total_score += 0.20
        elif sums_ok:
            print(f"PASS: Component 3 — Row 9 has SUM formulas but bold may be missing (0.12 pts)")
            total_score += 0.12
        elif sum_formulas_ok >= 3:
            print(f"PARTIAL: Component 3 — Row 9 has {sum_formulas_ok}/5 SUM formulas (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — Row 9 totals incomplete: {sum_formulas_ok}/5 SUM formulas, {bold_count}/6 bold cells (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Currency number format on F2:J8 (0.15 points)
    # All cells in F2:J8 should have currency-like number format
    try:
        currency_formats = 0
        total_cells = 0
        currency_indicators = ['$', '#,##0', '0.00', 'Currency']

        for row in range(2, 9):  # rows 2-8
            for col in range(6, 11):  # F to J
                total_cells += 1
                fmt = ws.cell(row=row, column=col).number_format
                if fmt and any(ind in fmt for ind in currency_indicators):
                    currency_formats += 1

        currency_ratio = currency_formats / total_cells if total_cells > 0 else 0
        if currency_ratio >= 0.90:
            print(f"PASS: Component 4 — {currency_formats}/{total_cells} cells have currency format (0.15 pts)")
            total_score += 0.15
        elif currency_ratio >= 0.60:
            partial = 0.08
            print(f"PARTIAL: Component 4 — {currency_formats}/{total_cells} cells have currency format ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {currency_formats}/{total_cells} cells have currency format (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Conditional formatting on H2:H8 and I2:I8 (0.10 points)
    # positive values -> green fill, negative values -> red fill
    try:
        cf_ranges = []
        for cf_range in ws.conditional_formatting:
            range_str = str(cf_range)
            cf_ranges.append(range_str)

        # Check if conditional formatting covers H2:H8 and I2:I8
        h_has_cf = any('H2' in r and 'H8' in r for r in cf_ranges) or any('H' in r for r in cf_ranges)
        i_has_cf = any('I2' in r and 'I8' in r for r in cf_ranges) or any('I' in r for r in cf_ranges)

        # More detailed check: verify green/red rules
        green_rules = 0
        red_rules = 0
        for cf_range in ws.conditional_formatting:
            range_str = str(cf_range)
            if 'H' in range_str or 'I' in range_str:
                for rule in cf_range.rules:
                    try:
                        if rule.dxf and rule.dxf.fill:
                            fill_color = rule.dxf.fill.fgColor.rgb
                            if rule.type == 'cellIs':
                                op = getattr(rule, 'operator', '')
                                if op == 'greaterThan':
                                    green_rules += 1
                                elif op == 'lessThan':
                                    red_rules += 1
                    except Exception:
                        pass

        if green_rules >= 2 and red_rules >= 2:
            print(f"PASS: Component 5 — Conditional formatting with green/red on H and I columns (0.10 pts)")
            total_score += 0.10
        elif (h_has_cf and i_has_cf) or green_rules >= 1 or red_rules >= 1:
            print(f"PARTIAL: Component 5 — Partial conditional formatting found (green_rules={green_rules}, red_rules={red_rules}) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — No conditional formatting on H or I columns. Ranges found: {cf_ranges} (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Bar chart present with 2 series referencing H and I columns (0.10 points)
    # A bar/column chart should show H9 and I9 side by side (volume vs price effect)
    try:
        charts = ws._charts
        if len(charts) == 0:
            print("FAIL: Component 6 — No charts found in PVM sheet (0.0 pts)")
        else:
            chart = charts[0]
            has_two_series = len(chart.series) >= 2

            # Check if series reference H and I columns
            series_refs = []
            for series in chart.series:
                try:
                    series_refs.append(str(series.val))
                except Exception:
                    pass

            h_referenced = any('H' in ref for ref in series_refs)
            i_referenced = any('I' in ref for ref in series_refs)

            if has_two_series and h_referenced and i_referenced:
                print(f"PASS: Component 6 — Bar chart with 2 series referencing H and I columns (0.10 pts)")
                total_score += 0.10
            elif has_two_series or (h_referenced and i_referenced):
                print(f"PARTIAL: Component 6 — Chart found but incomplete (series={len(chart.series)}, H_ref={h_referenced}, I_ref={i_referenced}) (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 6 — Chart found but does not properly reference H and I (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
