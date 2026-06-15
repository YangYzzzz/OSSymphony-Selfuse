"""
Initial Setup: Set up data validation on cell D2 for discount rate field
Task ID: calc_nrv_048
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_048'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    # Headers
    headers = ['Product', 'Price', 'Qty', 'Discount Rate']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Product data - realistic entries
    data = [
        ['Wireless Mouse',         29.99,  150, None],   # D2 empty - task target
        ['Mechanical Keyboard',    89.50,   75, 0.15],
        ['USB-C Hub',              45.00,  200, 0.10],
        ['27\" Monitor',          349.99,   40, 0.05],
        ['Laptop Stand',           34.95,  120, 0.20],
        ['Webcam HD 1080p',        59.99,   90, 0.12],
        ['Noise-Cancelling Headphones', 199.00, 55, 0.08],
        ['External SSD 1TB',      109.95,   65, 0.18],
        ['Desk Lamp LED',          24.50,  180, 0.25],
        ['Ergonomic Chair',       449.00,   30, 0.07],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Format price column as currency
    for r in range(2, 12):
        ws.cell(row=r, column=2).number_format = '$#,##0.00'

    # Format discount rate column as percentage where values exist
    for r in range(3, 12):
        ws.cell(row=r, column=4).number_format = '0.00%'

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 15

    # NO data validation on D2 - that is the task

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
