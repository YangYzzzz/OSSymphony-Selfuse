"""
Reward Script: Sales Rep Scorecard with Weighted KPI Scoring
Task ID: calc_sales_061
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Individual KPI score formulas (B2:E4) — MIN(Actual/Target, 1.5)
  Component 2 (0.3): Weighted score formulas (F2:F4) — B*0.4+C*0.2+D*0.2+E*0.2
  Component 3 (0.3): Rating formulas (G2:G4) — IF-based Exceeds/Meets/Below
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_061'


def normalize_formula(f):
    """Normalize a formula string for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ""
    return f.upper().replace(" ", "")


def check_kpi_score_formula(formula, row, kpi_col_actual, kpi_col_target):
    """
    Check if a formula computes MIN(Actual/Target, 1.5) for a given KPI.
    The formula should reference KPIs sheet columns for actual and target.
    Accepts variants like =MIN(KPIs!C2/KPIs!B2,1.5) or =MIN(C2/B2,1.5) etc.
    Also accepts direct numeric values matching ground truth.
    """
    if formula is None:
        return False
    norm = normalize_formula(formula)
    if not norm.startswith("="):
        return False

    # Check it contains MIN and 1.5 (the cap)
    if "MIN(" not in norm:
        return False
    if "1.5" not in norm:
        return False

    # Check it references actual/target columns for this row
    # KPI actual columns: C (revenue), E (logo), G (pipeline), I (activity)
    # KPI target columns: B (revenue), D (logo), F (pipeline), H (activity)
    actual_ref = f"{kpi_col_actual}{row}"
    target_ref = f"{kpi_col_target}{row}"

    # Could be prefixed with "KPIS!" or not
    has_actual = actual_ref in norm or f"KPIS!{actual_ref}" in norm
    has_target = target_ref in norm or f"KPIS!{target_ref}" in norm

    # Also check for division
    has_division = "/" in norm

    return has_actual and has_target and has_division


def check_weighted_formula(formula, row):
    """
    Check if formula computes weighted score: B*0.4+C*0.2+D*0.2+E*0.2
    Accepts various orderings and formats.
    """
    if formula is None:
        return False
    norm = normalize_formula(formula)
    if not norm.startswith("="):
        return False

    # Must reference B, C, D, E columns for this row with weights 0.4, 0.2, 0.2, 0.2
    r = str(row)

    # Check that B<row> appears with 0.4 weight
    has_b_04 = bool(re.search(rf'B{r}\*0\.4|0\.4\*B{r}', norm))
    # Check that C<row>, D<row>, E<row> each appear with 0.2 weight
    has_c_02 = bool(re.search(rf'C{r}\*0\.2|0\.2\*C{r}', norm))
    has_d_02 = bool(re.search(rf'D{r}\*0\.2|0\.2\*D{r}', norm))
    has_e_02 = bool(re.search(rf'E{r}\*0\.2|0\.2\*E{r}', norm))

    return has_b_04 and has_c_02 and has_d_02 and has_e_02


def check_rating_formula(formula, row):
    """
    Check if formula computes rating: IF(F>=1.1,"Exceeds",IF(F>=0.9,"Meets","Below"))
    Accepts various equivalent IF nesting patterns.
    """
    if formula is None:
        return False
    norm = normalize_formula(formula)
    if not norm.startswith("="):
        return False

    # Must contain IF, reference F<row>, and include the three rating labels
    r = str(row)
    has_if = "IF(" in norm
    has_f_ref = f"F{r}" in norm
    has_exceeds = '"EXCEEDS"' in norm
    has_meets = '"MEETS"' in norm
    has_below = '"BELOW"' in norm

    # Must contain threshold values
    has_threshold_high = "1.1" in norm
    has_threshold_low = "0.9" in norm

    return (has_if and has_f_ref and has_exceeds and has_meets
            and has_below and has_threshold_high and has_threshold_low)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Scorecard sheet must exist
    if 'Scorecard' not in wb.sheetnames:
        print("FAIL: 'Scorecard' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Scorecard']

    # Mapping of KPI score columns to their KPIs sheet actual/target columns
    # B = Revenue: actual=C, target=B on KPIs
    # C = Logo: actual=E, target=D on KPIs
    # D = Pipeline: actual=G, target=F on KPIs
    # E = Activity: actual=I, target=H on KPIs
    kpi_mappings = {
        'B': ('C', 'B'),   # Revenue score: actual=C, target=B
        'C': ('E', 'D'),   # Logo score: actual=E, target=D
        'D': ('G', 'F'),   # Pipeline score: actual=G, target=F
        'E': ('I', 'H'),   # Activity score: actual=I, target=H
    }

    # Component 1: Individual KPI score formulas B2:E4 (0.4 points)
    # Each of 12 cells (4 cols x 3 rows) contributes equally
    try:
        kpi_pass_count = 0
        kpi_total = 12  # 4 KPIs x 3 reps
        for col_letter, (actual_col, target_col) in kpi_mappings.items():
            for row in range(2, 5):
                cell_ref = f"{col_letter}{row}"
                formula = ws[cell_ref].value
                if check_kpi_score_formula(formula, row, actual_col, target_col):
                    kpi_pass_count += 1
                else:
                    print(f"  FAIL: {cell_ref} — expected MIN(KPIs!{actual_col}{row}/KPIs!{target_col}{row},1.5), found: {formula}")

        if kpi_pass_count == kpi_total:
            print(f"PASS: Component 1 — All 12 KPI score formulas correct ({kpi_pass_count}/{kpi_total}) (0.4 pts)")
            total_score += 0.4
        elif kpi_pass_count > 0:
            kpi_score = round(0.4 * (kpi_pass_count / kpi_total), 4)
            print(f"PARTIAL: Component 1 — {kpi_pass_count}/{kpi_total} KPI score formulas correct ({kpi_score:.2f} pts)")
            total_score += kpi_score
        else:
            print(f"FAIL: Component 1 — No KPI score formulas found (0/{kpi_total})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Weighted score formulas F2:F4 (0.3 points)
    # Each of 3 cells contributes equally
    try:
        weighted_pass_count = 0
        weighted_total = 3
        for row in range(2, 5):
            cell_ref = f"F{row}"
            formula = ws[cell_ref].value
            if check_weighted_formula(formula, row):
                weighted_pass_count += 1
            else:
                print(f"  FAIL: {cell_ref} — expected B{row}*0.4+C{row}*0.2+D{row}*0.2+E{row}*0.2, found: {formula}")

        if weighted_pass_count == weighted_total:
            print(f"PASS: Component 2 — All 3 weighted score formulas correct ({weighted_pass_count}/{weighted_total}) (0.3 pts)")
            total_score += 0.3
        elif weighted_pass_count > 0:
            weighted_score = round(0.3 * (weighted_pass_count / weighted_total), 4)
            print(f"PARTIAL: Component 2 — {weighted_pass_count}/{weighted_total} weighted score formulas correct ({weighted_score:.2f} pts)")
            total_score += weighted_score
        else:
            print(f"FAIL: Component 2 — No weighted score formulas found (0/{weighted_total})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Rating formulas G2:G4 (0.3 points)
    # Each of 3 cells contributes equally
    try:
        rating_pass_count = 0
        rating_total = 3
        for row in range(2, 5):
            cell_ref = f"G{row}"
            formula = ws[cell_ref].value
            if check_rating_formula(formula, row):
                rating_pass_count += 1
            else:
                print(f"  FAIL: {cell_ref} — expected IF(F{row}>=1.1,\"Exceeds\",IF(F{row}>=0.9,\"Meets\",\"Below\")), found: {formula}")

        if rating_pass_count == rating_total:
            print(f"PASS: Component 3 — All 3 rating formulas correct ({rating_pass_count}/{rating_total}) (0.3 pts)")
            total_score += 0.3
        elif rating_pass_count > 0:
            rating_score = round(0.3 * (rating_pass_count / rating_total), 4)
            print(f"PARTIAL: Component 3 — {rating_pass_count}/{rating_total} rating formulas correct ({rating_score:.2f} pts)")
            total_score += rating_score
        else:
            print(f"FAIL: Component 3 — No rating formulas found (0/{rating_total})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
