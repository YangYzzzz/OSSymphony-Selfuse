"""
Initial Setup: Sales rep scorecard with KPIs data and empty scorecard template
Task ID: calc_sales_061
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_061'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: KPIs ---
    ws_kpi = wb.active
    ws_kpi.title = 'KPIs'

    # Headers
    headers_kpi = [
        'Rep', 'Revenue Target', 'Revenue Actual',
        'New Logo Target', 'New Logo Actual',
        'Pipeline Target', 'Pipeline Actual',
        'Activity Target', 'Activity Actual'
    ]
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(headers_kpi, 1):
        cell = ws_kpi.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Data rows
    kpi_data = [
        ['Alice', 500000, 550000, 10, 12, 1500000, 1800000, 100, 95],
        ['Bob',   500000, 420000, 10,  8, 1500000, 1200000, 100, 110],
        ['Carol', 500000, 500000, 10, 10, 1500000, 1500000, 100, 100],
    ]
    for r, row_data in enumerate(kpi_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_kpi.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 1:
                cell.font = Font(bold=True)
            elif c in (2, 3, 6, 7):
                cell.number_format = '$#,##0'
            cell.alignment = Alignment(horizontal="center")

    # Adjust column widths
    col_widths_kpi = [10, 16, 16, 16, 16, 16, 16, 16, 16]
    for i, w in enumerate(col_widths_kpi):
        ws_kpi.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = w

    ws_kpi.freeze_panes = "A2"

    # --- Sheet 2: Scorecard (empty template - task is to fill formulas) ---
    ws_sc = wb.create_sheet('Scorecard')

    headers_sc = [
        'Rep', 'Rev Score', 'Logo Score', 'Pipeline Score',
        'Activity Score', 'Weighted Score', 'Rating'
    ]
    for col, h in enumerate(headers_sc, 1):
        cell = ws_sc.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Rep names only - leave score columns empty (that's the task)
    reps = ['Alice', 'Bob', 'Carol']
    for r, name in enumerate(reps, 2):
        cell = ws_sc.cell(row=r, column=1, value=name)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        # Add borders for empty cells too
        for c in range(2, 8):
            cell = ws_sc.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # Adjust column widths
    col_widths_sc = [10, 14, 14, 16, 16, 16, 12]
    for i, w in enumerate(col_widths_sc):
        ws_sc.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = w

    ws_sc.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
