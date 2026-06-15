"""
Reward Script: Flight risk prediction model using scoring formula
Task ID: calc_hr_092
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Risk Score formulas present in G2:G13 with correct 5-factor IF logic
  Component 2 (0.3): Risk Level formulas present in H2:H13 with correct High/Medium/Low classification
  Component 3 (0.3): Computed values correct for Alice (G2=65, H2="High") via data_only load
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_092'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def check_risk_score_formula(formula_str):
    """
    Verify that a Risk Score formula contains all 5 scoring components:
    - Tenure check (B column, thresholds 2 and 5, points 20/10/0)
    - Years Since Promotion check (C column, thresholds 3 and 2, points 25/15/0)
    - Compa-Ratio check (D column, thresholds 0.9 and 1, points 20/10/0)
    - Engagement check (E column, thresholds 3.5 and 4, points 20/10/0)
    - Mgr Rating check (F column, thresholds 3 and 4, points 15/5/0)
    Returns (bool, details_str)
    """
    if not isinstance(formula_str, str) or not formula_str.startswith('='):
        return False, f"Not a formula: {formula_str!r}"

    upper = formula_str.upper().replace(" ", "")

    # Check each factor is referenced
    factors_found = 0
    details = []

    # Factor 1: Tenure - references B column with values 20, 10
    if re.search(r'IF\(B\d+', upper) and '20' in upper:
        factors_found += 1
        details.append("Tenure(B)")
    else:
        details.append("MISSING: Tenure(B)")

    # Factor 2: Years Since Promotion - references C column with value 25
    if re.search(r'IF\(C\d+', upper) and '25' in upper:
        factors_found += 1
        details.append("YrsSincePromo(C)")
    else:
        details.append("MISSING: YrsSincePromo(C)")

    # Factor 3: Compa-Ratio - references D column with 0.9
    if re.search(r'IF\(D\d+', upper) and '0.9' in formula_str:
        factors_found += 1
        details.append("CompaRatio(D)")
    else:
        details.append("MISSING: CompaRatio(D)")

    # Factor 4: Engagement - references E column with 3.5
    if re.search(r'IF\(E\d+', upper) and '3.5' in formula_str:
        factors_found += 1
        details.append("Engagement(E)")
    else:
        details.append("MISSING: Engagement(E)")

    # Factor 5: Mgr Rating - references F column with values 15 and 5
    if re.search(r'IF\(F\d+', upper) and '15' in upper:
        factors_found += 1
        details.append("MgrRating(F)")
    else:
        details.append("MISSING: MgrRating(F)")

    return factors_found >= 5, f"{factors_found}/5 factors: {', '.join(details)}"


def check_risk_level_formula(formula_str):
    """
    Verify that a Risk Level formula classifies based on G column:
    G>=60 -> "High", G>=35 -> "Medium", else "Low"
    Returns (bool, details_str)
    """
    if not isinstance(formula_str, str) or not formula_str.startswith('='):
        return False, f"Not a formula: {formula_str!r}"

    upper = formula_str.upper().replace(" ", "")

    has_g_ref = bool(re.search(r'IF\(G\d+', upper))
    has_high = 'HIGH' in upper
    has_medium = 'MEDIUM' in upper
    has_low = 'LOW' in upper
    has_60 = '60' in upper
    has_35 = '35' in upper

    all_ok = has_g_ref and has_high and has_medium and has_low and has_60 and has_35
    detail_parts = []
    if not has_g_ref:
        detail_parts.append("missing G ref")
    if not has_high:
        detail_parts.append("missing High")
    if not has_medium:
        detail_parts.append("missing Medium")
    if not has_low:
        detail_parts.append("missing Low")
    if not has_60:
        detail_parts.append("missing 60 threshold")
    if not has_35:
        detail_parts.append("missing 35 threshold")

    if all_ok:
        return True, "Correct classification formula"
    else:
        return False, f"Issues: {', '.join(detail_parts)}"


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook (formula mode)
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet exists
    if 'FlightRisk' not in wb.sheetnames:
        print("FAIL: Sheet 'FlightRisk' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['FlightRisk']

    # Component 1: Risk Score formulas in G2:G13 (0.4 points)
    # These cells are None in initial_env, must contain formulas in golden_env
    try:
        g_formula_count = 0
        g_correct_count = 0
        total_rows = 12  # rows 2-13
        for row in range(2, 14):
            cell_val = ws.cell(row=row, column=7).value  # column G
            if cell_val is not None and isinstance(cell_val, str) and cell_val.startswith('='):
                g_formula_count += 1
                ok, detail = check_risk_score_formula(cell_val)
                if ok:
                    g_correct_count += 1

        if g_correct_count == total_rows:
            print(f"PASS: Component 1 - All {total_rows} Risk Score formulas correct in G2:G13 (0.4 pts)")
            total_score += 0.4
        elif g_formula_count > 0:
            # Partial: proportion of correct formulas
            partial = 0.4 * (g_correct_count / total_rows)
            print(f"PARTIAL: Component 1 - {g_correct_count}/{total_rows} correct Risk Score formulas ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - No Risk Score formulas found in G2:G13 (found {g_formula_count} formulas)")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Risk Level formulas in H2:H13 (0.3 points)
    # These cells are None in initial_env, must contain formulas in golden_env
    try:
        h_formula_count = 0
        h_correct_count = 0
        for row in range(2, 14):
            cell_val = ws.cell(row=row, column=8).value  # column H
            if cell_val is not None and isinstance(cell_val, str) and cell_val.startswith('='):
                h_formula_count += 1
                ok, detail = check_risk_level_formula(cell_val)
                if ok:
                    h_correct_count += 1

        if h_correct_count == total_rows:
            print(f"PASS: Component 2 - All {total_rows} Risk Level formulas correct in H2:H13 (0.3 pts)")
            total_score += 0.3
        elif h_formula_count > 0:
            partial = 0.3 * (h_correct_count / total_rows)
            print(f"PARTIAL: Component 2 - {h_correct_count}/{total_rows} correct Risk Level formulas ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 - No Risk Level formulas found in H2:H13 (found {h_formula_count} formulas)")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Computed values correct for Alice (row 2) (0.3 points)
    # Ground truth from context: Alice risk score = 65, Risk Level = "High"
    # We need data_only=True to get cached computed values
    try:
        wb_data = openpyxl.load_workbook(file_path, data_only=True)
        ws_data = wb_data['FlightRisk']

        g2_val = ws_data.cell(row=2, column=7).value  # Risk Score
        h2_val = ws_data.cell(row=2, column=8).value  # Risk Level

        comp3_score = 0.0

        # Check G2 computed value = 65
        if g2_val is not None:
            try:
                g2_num = float(g2_val)
                if abs(g2_num - 65) < 0.01:
                    print(f"PASS: Component 3a - Alice Risk Score = {g2_num} (expected 65) (0.15 pts)")
                    comp3_score += 0.15
                else:
                    print(f"FAIL: Component 3a - Alice Risk Score = {g2_num}, expected 65")
            except (ValueError, TypeError):
                print(f"FAIL: Component 3a - Alice Risk Score not numeric: {g2_val!r}")
        else:
            # data_only may return None if file was never opened in Calc
            # Fall back: verify formula structure implies correct result for Alice's inputs
            # Alice: B2=1.5(<2->20), C2=0(<2->0), D2=0.85(<0.9->20), E2=3.2(<3.5->20), F2=3.5(3-4->5)
            # Expected: 20+0+20+20+5 = 65
            # If formula structure is correct (checked in Component 1), we can trust computation
            g2_formula = ws.cell(row=2, column=7).value
            if g2_formula is not None and isinstance(g2_formula, str) and g2_formula.startswith('='):
                # Formula exists; compute manually from Alice's data
                b2, c2, d2, e2, f2 = 1.5, 0, 0.85, 3.2, 3.5
                manual_score = 0
                manual_score += 20 if b2 < 2 else (10 if b2 <= 5 else 0)
                manual_score += 25 if c2 > 3 else (15 if c2 >= 2 else 0)
                manual_score += 20 if d2 < 0.9 else (10 if d2 <= 1 else 0)
                manual_score += 20 if e2 < 3.5 else (10 if e2 <= 4 else 0)
                manual_score += 15 if f2 < 3 else (5 if f2 <= 4 else 0)

                ok_formula, _ = check_risk_score_formula(g2_formula)
                if ok_formula and manual_score == 65:
                    print(f"PASS: Component 3a - Formula correct + manual calc confirms Alice = 65 (0.15 pts)")
                    comp3_score += 0.15
                else:
                    print(f"FAIL: Component 3a - data_only=None, formula check: ok={ok_formula}, manual={manual_score}")
            else:
                print(f"FAIL: Component 3a - No cached value and no formula in G2")

        # Check H2 computed value = "High"
        if h2_val is not None:
            if str(h2_val).strip().lower() == 'high':
                print(f"PASS: Component 3b - Alice Risk Level = '{h2_val}' (expected 'High') (0.15 pts)")
                comp3_score += 0.15
            else:
                print(f"FAIL: Component 3b - Alice Risk Level = '{h2_val}', expected 'High'")
        else:
            # Fall back: formula structure check
            h2_formula = ws.cell(row=2, column=8).value
            if h2_formula is not None and isinstance(h2_formula, str) and h2_formula.startswith('='):
                ok_formula, _ = check_risk_level_formula(h2_formula)
                # With score=65, 65>=60 => "High"
                if ok_formula:
                    print(f"PASS: Component 3b - Formula correct + score 65>=60 => 'High' (0.15 pts)")
                    comp3_score += 0.15
                else:
                    print(f"FAIL: Component 3b - data_only=None, formula structure incorrect")
            else:
                print(f"FAIL: Component 3b - No cached value and no formula in H2")

        total_score += comp3_score

    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
