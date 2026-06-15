"""
Initial Setup: Flight risk prediction model spreadsheet
Task ID: calc_hr_092
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_092'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'FlightRisk'

    # Headers
    headers = [
        'Employee', 'Tenure (yrs)', 'Years Since Promotion',
        'Compa-Ratio', 'Engagement', 'Mgr Rating', 'Risk Score', 'Risk Level'
    ]
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # Employee data - realistic HR data
    # Columns: Employee, Tenure(yrs), YearsSincePromo, Compa-Ratio, Engagement, MgrRating
    # G (Risk Score) and H (Risk Level) are LEFT EMPTY - that's the task
    employees = [
        ['Alice',            1.5,  0,   0.85, 3.2, 3.5],
        ['Brian Torres',     3.2,  2.5, 0.95, 3.8, 4.2],
        ['Catherine Nguyen', 7.1,  1.0, 1.05, 4.3, 4.5],
        ['Derek Washington', 0.8,  0,   0.88, 2.9, 2.8],
        ['Elena Petrov',     4.5,  3.5, 0.92, 3.6, 3.0],
        ['Faisal Ahmed',     6.3,  4.0, 1.10, 4.1, 4.8],
        ['Grace Kim',        2.0,  1.5, 0.78, 3.3, 3.2],
        ['Hector Ramirez',   1.2,  0,   1.02, 4.5, 4.1],
        ['Irene Johansson',  5.8,  2.0, 0.89, 3.0, 2.5],
        ['James Chen',       3.7,  3.0, 1.15, 4.6, 4.7],
        ['Karen Okonkwo',    0.5,  0,   0.82, 2.7, 3.8],
        ['Luis Fernandez',   8.2,  5.0, 1.08, 3.9, 3.5],
    ]

    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r, row_data in enumerate(employees, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = data_border
            if c == 1:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="center")
                if c == 4:  # Compa-Ratio
                    cell.number_format = '0.00'
                elif c == 5:  # Engagement
                    cell.number_format = '0.0'
                elif c == 2:  # Tenure
                    cell.number_format = '0.0'
                elif c == 6:  # Mgr Rating
                    cell.number_format = '0.0'

        # Add borders on empty G and H columns too
        for c in [7, 8]:
            cell = ws.cell(row=r, column=c)
            cell.border = data_border
            cell.alignment = Alignment(horizontal="center")

    # Set column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 13
    ws.column_dimensions['G'].width = 13
    ws.column_dimensions['H'].width = 13

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
