"""
Initial Setup: HR Shift Schedule spreadsheet (no conditional formatting)
Task ID: osworld_calc_conditional_format_weekday_010
Domain: libreoffice_calc

Creates an HR shift schedule with realistic data.
NO conditional formatting is applied — the agent must add it.
"""

import os
import shlex
import subprocess
import time
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_conditional_format_weekday_010'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shift Schedule"

    # Headers
    headers = ['Shift Date', 'Employee Name', 'Shift Start', 'Shift End', 'Department']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # Realistic shift data covering multiple weeks (mix of weekdays and weekends)
    # Starting from Monday 2025-03-03
    base_date = date(2025, 3, 3)  # Monday

    employees = [
        ('Sarah Chen', 'Nursing'),
        ('Marcus Johnson', 'Emergency'),
        ('Priya Patel', 'Radiology'),
        ('David O\'Brien', 'Pharmacy'),
        ('Linda Torres', 'Nursing'),
        ('James Kim', 'Emergency'),
        ('Aisha Williams', 'Surgery'),
        ('Robert Nguyen', 'Radiology'),
        ('Emily Rodriguez', 'Nursing'),
        ('Michael Okafor', 'Pharmacy'),
        ('Fatima Hassan', 'Surgery'),
        ('Carlos Mendez', 'Emergency'),
        ('Natalie Brooks', 'Nursing'),
        ('Steven Park', 'Radiology'),
        ('Diana Scott', 'Surgery'),
    ]

    shifts = [
        ('07:00', '15:00'),
        ('15:00', '23:00'),
        ('23:00', '07:00'),
        ('08:00', '16:00'),
        ('12:00', '20:00'),
    ]

    row = 2
    emp_idx = 0
    for day_offset in range(15):  # 15 days of data
        current_date = base_date + timedelta(days=day_offset)
        date_str = current_date.strftime('%Y-%m-%d')

        # Number of employees per day: 3-4
        num_shifts = 3 if day_offset % 3 != 0 else 4
        for _ in range(num_shifts):
            emp_name, dept = employees[emp_idx % len(employees)]
            shift_start, shift_end = shifts[(emp_idx + day_offset) % len(shifts)]
            ws.cell(row=row, column=1, value=date_str)
            ws.cell(row=row, column=2, value=emp_name)
            ws.cell(row=row, column=3, value=shift_start)
            ws.cell(row=row, column=4, value=shift_end)
            ws.cell(row=row, column=5, value=dept)
            emp_idx += 1
            row += 1

    # Column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
