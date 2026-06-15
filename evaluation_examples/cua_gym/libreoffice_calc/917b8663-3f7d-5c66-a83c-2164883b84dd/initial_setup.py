"""
Initial Setup: Create a spreadsheet with Records sheet containing record names and numbers.
Task ID: calc_lf_067
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_067'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Records ---
    ws = wb.active
    ws.title = 'Records'

    # Headers
    ws.cell(row=1, column=1, value='Record')
    ws.cell(row=1, column=2, value='Number')

    # Data rows - numeric values, NO custom format applied
    data = [
        ['First', 42],
        ['Second', 7],
        ['Third', 1358],
    ]
    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        ws.cell(row=r, column=2, value=row_data[1])

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
