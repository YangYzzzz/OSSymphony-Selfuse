"""
Reward Script: Apply custom number format "ID-"00000 to cells B2:B4
Task ID: calc_lf_067
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): B2 has correct number format and numeric value 42
  Component 2 (0.3): B3 has correct number format and numeric value 7
  Component 3 (0.3): B4 has correct number format and numeric value 1358
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_067'
EXPECTED_FORMAT = '"ID-"00000'

# Expected cell values and weights
CHECKS = [
    # (cell_coord, expected_value, weight, label)
    ("B2", 42, 0.4, "B2: format + numeric value 42"),
    ("B3", 7, 0.3, "B3: format + numeric value 7"),
    ("B4", 1358, 0.3, "B4: format + numeric value 1358"),
]


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Records' sheet must exist
    if 'Records' not in wb.sheetnames:
        print(f"FAIL: Sheet 'Records' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Records']

    for idx, (coord, expected_val, weight, label) in enumerate(CHECKS, 1):
        # Component N: Check number format is "ID-"00000 AND value is numeric and correct
        try:
            cell = ws[coord]
            fmt = cell.number_format
            val = cell.value

            fmt_ok = (fmt == EXPECTED_FORMAT)
            val_numeric = isinstance(val, (int, float))
            val_correct = (val_numeric and val == expected_val)

            if fmt_ok and val_correct:
                print(f"PASS: Component {idx} -- {label}: format={fmt!r}, value={val!r} ({weight} pts)")
                total_score += weight
            else:
                reasons = []
                if not fmt_ok:
                    reasons.append(f"format={fmt!r}, expected={EXPECTED_FORMAT!r}")
                if not val_numeric:
                    reasons.append(f"value is not numeric: {type(val).__name__}={val!r}")
                elif not val_correct:
                    reasons.append(f"value={val!r}, expected={expected_val!r}")
                print(f"FAIL: Component {idx} -- {label}: {'; '.join(reasons)}")
        except Exception as e:
            print(f"ERROR: Component {idx} -- {label}: {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
