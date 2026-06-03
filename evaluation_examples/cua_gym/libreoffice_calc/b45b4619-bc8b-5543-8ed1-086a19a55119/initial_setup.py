"""
Initial Setup: Build a conversion funnel analysis for inbound marketing process
Task ID: calc_sales_marketing_funnel_029
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_marketing_funnel_029'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: FunnelData ---
    ws = wb.active
    ws.title = 'FunnelData'

    # Headers in row 1
    headers = ['Stage', 'Count', 'Stage Conversion %', 'Overall Conversion %']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Funnel stage data (Columns C and D intentionally left empty - task requires filling them)
    funnel_data = [
        ('Visitors',      125000),
        ('Leads',           8750),
        ('MQLs',            2100),
        ('SQLs',             840),
        ('Opportunities',    420),
        ('Customers',        168),
    ]

    for r, (stage, count) in enumerate(funnel_data, 2):
        ws.cell(row=r, column=1, value=stage)
        ws.cell(row=r, column=2, value=count)
        # Columns C (Stage Conversion %) and D (Overall Conversion %) are intentionally empty
        # The agent must fill these in as part of the task

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
