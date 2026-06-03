"""
Reward Script: Build a conversion funnel analysis for inbound marketing process
Task ID: calc_sales_marketing_funnel_029
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.30): Stage conversion formulas in C3:C7 (stage-to-stage ratios)
  - Component 2 (0.30): Overall conversion formulas in D2:D7 (ratios vs Visitors)
  - Component 3 (0.20): Percentage number format applied to C2:C7 and D2:D7
  - Component 4 (0.20): Bar chart present with title 'Marketing Conversion Funnel'
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_marketing_funnel_029'

# Expected stage-to-stage conversion formulas in column C
# C3=B3/B2, C4=B4/B3, C5=B5/B4, C6=B6/B5, C7=B7/B6
EXPECTED_STAGE_FORMULAS = {
    3: ('=B3/B2', 'Leads/Visitors'),
    4: ('=B4/B3', 'MQLs/Leads'),
    5: ('=B5/B4', 'SQLs/MQLs'),
    6: ('=B6/B5', 'Opportunities/SQLs'),
    7: ('=B7/B6', 'Customers/Opportunities'),
}

# Expected overall conversion formulas in column D
# D3=B3/$B$2, D4=B4/$B$2, D5=B5/$B$2, D6=B6/$B$2, D7=B7/$B$2
EXPECTED_OVERALL_FORMULAS = {
    3: ('=B3/$B$2', 'Leads overall'),
    4: ('=B4/$B$2', 'MQLs overall'),
    5: ('=B5/$B$2', 'SQLs overall'),
    6: ('=B6/$B$2', 'Opportunities overall'),
    7: ('=B7/$B$2', 'Customers overall'),
}


def check_formula_match(actual, expected):
    """Check if actual formula matches expected, allowing for $ sign variations."""
    if not actual or not isinstance(actual, str):
        return False
    # Normalize both: uppercase, remove spaces
    actual_norm = actual.upper().replace(' ', '')
    expected_norm = expected.upper().replace(' ', '')
    # Check with and without $ anchors
    actual_no_dollar = actual_norm.replace('$', '')
    expected_no_dollar = expected_norm.replace('$', '')
    return actual_norm == expected_norm or actual_no_dollar == expected_no_dollar


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sheet 'FunnelData' must exist
    if 'FunnelData' not in wb.sheetnames:
        print("CRITICAL: Sheet 'FunnelData' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['FunnelData']

    # -------------------------------------------------------------------------
    # Component 1: Stage conversion formulas in C3:C7 (0.30 points)
    # Each row should have formula =B{n}/B{n-1} (stage-to-stage ratio)
    # C2 should be 1 (100%) as baseline
    # -------------------------------------------------------------------------
    try:
        stage_passes = 0

        for row, (expected_formula, desc) in EXPECTED_STAGE_FORMULAS.items():
            actual = ws.cell(row=row, column=3).value
            if check_formula_match(str(actual) if actual is not None else '', expected_formula):
                stage_passes += 1
                print(f"PASS: C{row} stage conversion formula '{actual}' ({desc})")
            else:
                print(f"FAIL: C{row} stage conversion — expected '{expected_formula}', found {repr(actual)}")

        # Also check C2 is set to some form of 100% (value 1)
        c2_val = ws.cell(row=2, column=3).value
        c2_is_base = (c2_val in (1, 1.0)) or (isinstance(c2_val, str) and c2_val.strip() in ('1', '100%'))
        if c2_is_base:
            print(f"PASS: C2 base stage conversion = {repr(c2_val)} (100%)")
        else:
            print(f"FAIL: C2 base stage conversion — expected 1 (100%), found {repr(c2_val)}")

        # Scoring based on how many of the 5 stage formulas are correct
        if stage_passes == 5:
            component1_score = 0.30
            print(f"PASS: Component 1 — All 5 stage conversion formulas correct (0.30 pts)")
        elif stage_passes >= 3:
            component1_score = 0.15
            print(f"PARTIAL: Component 1 — {stage_passes}/5 stage conversion formulas correct (0.15 pts)")
        elif stage_passes >= 1:
            component1_score = 0.06
            print(f"PARTIAL: Component 1 — {stage_passes}/5 stage conversion formulas correct (0.06 pts)")
        else:
            component1_score = 0.0
            print(f"FAIL: Component 1 — No stage conversion formulas found (0.0 pts)")

        total_score += component1_score

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Overall conversion formulas in D2:D7 (0.30 points)
    # D2 should be 1 (100%), D3:D7 should be =B{n}/$B$2 (ratio vs Visitors)
    # -------------------------------------------------------------------------
    try:
        overall_passes = 0

        for row, (expected_formula, desc) in EXPECTED_OVERALL_FORMULAS.items():
            actual = ws.cell(row=row, column=4).value
            if check_formula_match(str(actual) if actual is not None else '', expected_formula):
                overall_passes += 1
                print(f"PASS: D{row} overall conversion formula '{actual}' ({desc})")
            else:
                print(f"FAIL: D{row} overall conversion — expected '{expected_formula}', found {repr(actual)}")

        # Check D2 is 1 (100%) as base
        d2_val = ws.cell(row=2, column=4).value
        d2_is_base = (d2_val in (1, 1.0)) or (isinstance(d2_val, str) and d2_val.strip() in ('1', '100%'))
        if d2_is_base:
            print(f"PASS: D2 base overall conversion = {repr(d2_val)} (100%)")
        else:
            print(f"FAIL: D2 base overall conversion — expected 1 (100%), found {repr(d2_val)}")

        # Scoring based on how many of the 5 overall formulas are correct
        if overall_passes == 5:
            component2_score = 0.30
            print(f"PASS: Component 2 — All 5 overall conversion formulas correct (0.30 pts)")
        elif overall_passes >= 3:
            component2_score = 0.15
            print(f"PARTIAL: Component 2 — {overall_passes}/5 overall conversion formulas correct (0.15 pts)")
        elif overall_passes >= 1:
            component2_score = 0.06
            print(f"PARTIAL: Component 2 — {overall_passes}/5 overall conversion formulas correct (0.06 pts)")
        else:
            component2_score = 0.0
            print(f"FAIL: Component 2 — No overall conversion formulas found (0.0 pts)")

        total_score += component2_score

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Percentage number format on C2:C7 and D2:D7 (0.20 points)
    # Cells should have a percentage number format (containing '%')
    # -------------------------------------------------------------------------
    try:
        pct_format_passes = 0
        pct_cells = [
            (2, 3), (3, 3), (4, 3), (5, 3), (6, 3), (7, 3),  # C2:C7
            (2, 4), (3, 4), (4, 4), (5, 4), (6, 4), (7, 4),  # D2:D7
        ]

        for row, col in pct_cells:
            cell = ws.cell(row=row, column=col)
            fmt = cell.number_format or 'General'
            if '%' in fmt:
                pct_format_passes += 1
            else:
                from openpyxl.utils import get_column_letter
                coord = f"{get_column_letter(col)}{row}"
                print(f"FAIL: {coord} number format is {repr(fmt)}, expected percentage format")

        if pct_format_passes == 12:
            component3_score = 0.20
            print(f"PASS: Component 3 — All 12 cells have percentage number format (0.20 pts)")
        elif pct_format_passes >= 6:
            component3_score = 0.10
            print(f"PARTIAL: Component 3 — {pct_format_passes}/12 cells have percentage format (0.10 pts)")
        else:
            component3_score = 0.0
            print(f"FAIL: Component 3 — Only {pct_format_passes}/12 cells have percentage format (0.0 pts)")

        total_score += component3_score

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: Bar chart present with title 'Marketing Conversion Funnel' (0.20 points)
    # Task requires a funnel-style horizontal bar chart on the same sheet
    # -------------------------------------------------------------------------
    try:
        charts = ws._charts
        num_charts = len(charts)

        if num_charts == 0:
            print("FAIL: Component 4 — No charts found on FunnelData sheet (0.0 pts)")
            total_score += 0.0
        else:
            chart = charts[0]
            is_bar_type = type(chart).__name__ == 'BarChart'

            # Check chart title contains 'Marketing Conversion Funnel'
            chart_title_found = False
            try:
                if chart.title:
                    title_str = str(chart.title)
                    if 'Marketing Conversion Funnel' in title_str:
                        chart_title_found = True
                    # Also navigate rich text title structure
                    if hasattr(chart.title, 'tx') and chart.title.tx:
                        tx = chart.title.tx
                        if hasattr(tx, 'rich') and tx.rich:
                            for para in tx.rich.p:
                                for run in para.r:
                                    if 'Marketing Conversion Funnel' in run.t:
                                        chart_title_found = True
            except Exception as te:
                print(f"  WARNING: Could not read chart title: {te}")

            # Check chart has data referencing column B (Count values)
            chart_data_ref_ok = False
            try:
                if chart.series:
                    for ser in chart.series:
                        if hasattr(ser, 'val') and ser.val:
                            val_str = str(ser.val)
                            if '$B$' in val_str or 'FunnelData' in val_str:
                                chart_data_ref_ok = True
            except Exception as de:
                print(f"  WARNING: Could not check chart data reference: {de}")

            print(f"  Chart: type={type(chart).__name__}, is_bar={is_bar_type}, "
                  f"title_found={chart_title_found}, data_ref_ok={chart_data_ref_ok}")

            if chart_title_found and is_bar_type:
                component4_score = 0.20
                print(f"PASS: Component 4 — Horizontal bar chart with correct title found (0.20 pts)")
            elif chart_title_found or (is_bar_type and chart_data_ref_ok):
                component4_score = 0.10
                print(f"PARTIAL: Component 4 — Chart found but missing title or wrong type (0.10 pts)")
            elif num_charts > 0:
                component4_score = 0.05
                print(f"PARTIAL: Component 4 — Chart found but wrong type and missing title (0.05 pts)")
            else:
                component4_score = 0.0
                print(f"FAIL: Component 4 — No valid chart (0.0 pts)")

            total_score += component4_score

    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
