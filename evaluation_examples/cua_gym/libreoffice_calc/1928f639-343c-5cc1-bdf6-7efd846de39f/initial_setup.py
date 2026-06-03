"""
Initial Setup: Crop yield analysis spreadsheet for a farm
Task ID: calc_wf_085
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_085'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Fields ---
    ws = wb.active
    ws.title = 'Fields'

    # Headers
    headers = [
        'Field Name', 'Crop', 'Acres', 'Seed Cost', 'Fertilizer Cost',
        'Labor Hours', 'Labor Rate', 'Harvest (bushels/bales)', 'Market Price'
    ]
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # 15 fields of realistic farm data
    # Crops: Corn, Wheat, Soy, Cotton
    # Harvest units: bushels for Corn/Wheat/Soy, bales for Cotton
    data = [
        ['North Ridge',       'Corn',    120, 4800,  3600, 280, 18.50, 21600, 4.25],
        ['Sunrise Meadow',    'Wheat',    85, 2550,  2125, 195, 17.00, 4250,  5.80],
        ['Valley Bottom',     'Soy',      95, 3800,  2850, 210, 17.50, 4750,  10.50],
        ['Hilltop East',      'Cotton',   75, 3375,  2625, 240, 19.00, 825,   72.00],
        ['Creek Side',        'Corn',    150, 6000,  4500, 340, 18.50, 27000, 4.25],
        ['South Forty',       'Wheat',   110, 3300,  2750, 250, 17.00, 5500,  5.80],
        ['Pine Border',       'Soy',      60, 2400,  1800, 140, 17.50, 3000,  10.50],
        ['Windmill Field',    'Cotton',  100, 4500,  3500, 310, 19.00, 1100,  72.00],
        ['River Bend',        'Corn',     90, 3600,  2700, 205, 18.50, 16200, 4.25],
        ['Oak Grove',         'Wheat',    70, 2100,  1750, 160, 17.00, 3500,  5.80],
        ['Prairie West',      'Soy',     130, 5200,  3900, 290, 17.50, 6500,  10.50],
        ['Bluff Terrace',     'Cotton',   55, 2475,  1925, 175, 19.00, 605,   72.00],
        ['Lakeside North',    'Corn',    105, 4200,  3150, 240, 18.50, 18900, 4.25],
        ['Clover Patch',      'Wheat',    80, 2400,  2000, 185, 17.00, 4000,  5.80],
        ['Fenceline South',   'Soy',      65, 2600,  1950, 150, 17.50, 3250,  10.50],
    ]

    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    money_fmt = '$#,##0.00'
    number_fmt = '#,##0'
    rate_fmt = '$#,##0.00'

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = data_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # Format money columns
            if c in (4, 5):  # Seed Cost, Fertilizer Cost
                cell.number_format = money_fmt
            elif c == 7:  # Labor Rate
                cell.number_format = rate_fmt
            elif c in (3, 6, 8):  # Acres, Labor Hours, Harvest
                cell.number_format = number_fmt
            elif c == 9:  # Market Price
                cell.number_format = rate_fmt

    # Set column widths
    col_widths = {
        'A': 20, 'B': 12, 'C': 10, 'D': 14, 'E': 16,
        'F': 14, 'G': 14, 'H': 24, 'I': 14
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Row 1 height
    ws.row_dimensions[1].height = 30

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
