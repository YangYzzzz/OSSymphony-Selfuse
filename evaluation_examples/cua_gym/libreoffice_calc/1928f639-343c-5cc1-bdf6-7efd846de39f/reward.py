"""
Reward Script: Crop yield analysis spreadsheet
Task ID: calc_wf_085
Domain: libreoffice_calc
Scoring:
  Component 1: Calculated formula columns J-N (0.25)
  Component 2: Crop Type Summary with AVERAGEIF/SUMIF/COUNTIF (0.20)
  Component 3: Most Profitable Crop INDEX/MATCH formula (0.10)
  Component 4: At least 2 charts present (0.20)
  Component 5: Conditional formatting on yield column (0.15)
  Component 6: Weather Impact Notes section (0.10)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_085'


def persist_app_state(domain):
    """Save any unsaved LibreOffice state."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that 'Fields' sheet exists
    if 'Fields' not in wb.sheetnames:
        print("CRITICAL: 'Fields' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Fields']

    # Component 1: Calculated formula columns J-N for all 15 fields (0.25 points)
    # Initial file has only cols A-I. Golden adds J (Yield/Acre), K (Total Cost),
    # L (Revenue), M (Profit), N (Profit/Acre) with formulas.
    try:
        formula_count = 0
        expected_formulas = 15 * 5  # 15 rows x 5 columns (J-N)
        for row in range(2, 17):  # rows 2-16
            for col in range(10, 15):  # J=10, K=11, L=12, M=13, N=14
                val = ws.cell(row=row, column=col).value
                if val is not None and isinstance(val, str) and val.startswith('='):
                    formula_count += 1
        ratio = formula_count / expected_formulas
        if ratio >= 0.9:
            print(f"PASS: Component 1 - {formula_count}/{expected_formulas} formula cells in J-N (0.25 pts)")
            total_score += 0.25
        elif ratio >= 0.5:
            partial = 0.25 * ratio
            print(f"PARTIAL: Component 1 - {formula_count}/{expected_formulas} formula cells ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - Only {formula_count}/{expected_formulas} formula cells in J-N")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Crop Type Summary section with AVERAGEIF/SUMIF/COUNTIF (0.20 points)
    # Golden has summary starting around row 18-23 with crop types Corn, Wheat, Soy, Cotton
    # and AVERAGEIF for avg yield, AVERAGEIF for avg profit, SUMIF for total profit, COUNTIF for count
    try:
        summary_found = False
        crop_types_found = 0
        aggregate_formulas = 0
        crop_names = {'Corn', 'Wheat', 'Soy', 'Cotton'}

        # Search rows 17-30 for the summary section
        for row in range(17, 31):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val and str(cell_val).strip() in crop_names:
                crop_types_found += 1
                # Check columns B-E for aggregate formulas
                for col in range(2, 6):
                    v = ws.cell(row=row, column=col).value
                    if v and isinstance(v, str) and '=' in v:
                        upper_v = v.upper()
                        if any(f in upper_v for f in ['AVERAGEIF', 'SUMIF', 'COUNTIF']):
                            aggregate_formulas += 1

        if crop_types_found >= 4 and aggregate_formulas >= 8:
            print(f"PASS: Component 2 - Summary section: {crop_types_found} crop types, {aggregate_formulas} aggregate formulas (0.20 pts)")
            total_score += 0.20
        elif crop_types_found >= 2 and aggregate_formulas >= 4:
            print(f"PARTIAL: Component 2 - {crop_types_found} crops, {aggregate_formulas} formulas (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 - {crop_types_found} crop types found, {aggregate_formulas} aggregate formulas")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Most Profitable Crop formula using INDEX/MATCH (0.10 points)
    # Golden has this around row 25: =INDEX(A20:A23,MATCH(MAX(C20:C23),C20:C23,0))
    try:
        index_match_cells = [
            ws.cell(row=r, column=c).value
            for r in range(17, 35)
            for c in range(1, 15)
            if ws.cell(row=r, column=c).value and isinstance(ws.cell(row=r, column=c).value, str)
            and 'INDEX' in str(ws.cell(row=r, column=c).value).upper()
            and 'MATCH' in str(ws.cell(row=r, column=c).value).upper()
            and 'MAX' in str(ws.cell(row=r, column=c).value).upper()
        ]

        if len(index_match_cells) > 0:
            print(f"PASS: Component 3 - INDEX/MATCH/MAX formula for most profitable crop (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 - No INDEX/MATCH/MAX formula found for most profitable crop")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: At least 2 charts present (0.20 points)
    # Golden has: bar chart "Profit per Field" and grouped bar "Crop Comparison"
    try:
        charts = ws._charts
        num_charts = len(charts) if charts else 0

        if num_charts >= 2:
            print(f"PASS: Component 4 - {num_charts} charts found (0.20 pts)")
            total_score += 0.20
        elif num_charts == 1:
            print(f"PARTIAL: Component 4 - Only 1 chart found (0.10 pts)")
            total_score += 0.10
        else:
            # Also check other sheets for charts
            other_charts = 0
            for sn in wb.sheetnames:
                if sn != 'Fields':
                    other_ws = wb[sn]
                    other_charts += len(other_ws._charts) if other_ws._charts else 0
            total_charts = num_charts + other_charts
            if total_charts >= 2:
                print(f"PASS: Component 4 - {total_charts} charts found across sheets (0.20 pts)")
                total_score += 0.20
            elif total_charts == 1:
                print(f"PARTIAL: Component 4 - Only 1 chart found across sheets (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 4 - No charts found")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Conditional formatting on yield column for below-average yields (0.15 points)
    # Golden has CF on J2:J16 with expression rule using AVERAGEIF to identify below-average
    try:
        cf_rules = ws.conditional_formatting
        cf_on_yield = [
            cf for cf in cf_rules
            if 'J' in str(cf).upper()
            and any(rule.type in ('expression', 'cellIs') for rule in cf.rules)
        ]

        if len(cf_on_yield) > 0:
            print(f"PASS: Component 5 - Conditional formatting found on yield column (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 5 - No conditional formatting on yield column J")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    # Component 6: Weather Impact Notes section present (0.10 points)
    # Golden has "Weather Impact Notes" header and several note entries
    try:
        weather_header_row = None
        weather_notes_count = 0
        for row in range(17, 40):
            v = ws.cell(row=row, column=1).value
            if v and isinstance(v, str):
                if weather_header_row is None and 'weather' in v.lower() and 'note' in v.lower():
                    weather_header_row = row
                elif weather_header_row is not None and len(v.strip()) > 20:
                    weather_notes_count += 1

        if weather_header_row is not None and weather_notes_count >= 3:
            print(f"PASS: Component 6 - Weather Impact Notes with {weather_notes_count} entries (0.10 pts)")
            total_score += 0.10
        elif weather_header_row is not None:
            print(f"PARTIAL: Component 6 - Header found but only {weather_notes_count} notes (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 - No Weather Impact Notes section found")
    except Exception as e:
        print(f"ERROR: Component 6 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
