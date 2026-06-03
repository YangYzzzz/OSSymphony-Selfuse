"""
Initial Setup: Monthly revenue spreadsheet with partial formulas (D2 and E2 only).
Task ID: osworld_calc_formula_pattern_concat_005
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_formula_pattern_concat_005'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Revenue ---
    ws = wb.active
    ws.title = 'Revenue'

    # Headers in row 1
    headers = ['Month', 'Revenue', '', 'Running Total', 'Cumulative %', '']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic monthly revenue data (12 months)
    months_data = [
        ('January',   48320.50),
        ('February',  52140.75),
        ('March',     61870.25),
        ('April',     55490.00),
        ('May',       67230.80),
        ('June',      72450.60),
        ('July',      69180.40),
        ('August',    74620.90),
        ('September', 63940.25),
        ('October',   78350.00),
        ('November',  85670.15),
        ('December',  91430.70),
    ]

    for r, (month, revenue) in enumerate(months_data, 2):
        ws.cell(row=r, column=1, value=month)    # A: Month name
        ws.cell(row=r, column=2, value=revenue)  # B: Revenue
        # C column: left empty (unused)
        # D column: Running Total formula - ONLY D2 gets the formula
        # E column: Cumulative % formula - ONLY E2 gets the formula
        # F column: Label concatenation - ALL EMPTY in initial

    # D2 only: running total formula
    ws.cell(row=2, column=4, value='=SUM($B$2:B2)')

    # Total revenue for cumulative % denominator (sum of all 12 months)
    # E2 only: cumulative percentage formula
    ws.cell(row=2, column=5, value='=D2/SUM($B$2:$B$13)')

    # D3:D13 and E3:E13 are intentionally left EMPTY
    # F column (column 6) is entirely EMPTY

    # Column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 55

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
