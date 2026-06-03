"""
Reward Script: Extend running total/cumulative % formulas and create label concatenation in column F
Task ID: osworld_calc_formula_pattern_concat_005
Domain: libreoffice_calc
Scoring:
  - Component 1: Column D filled down for rows 3-13 with SUM($B$2:Bn) formulas (0.35 pts)
  - Component 2: Column E filled down for rows 3-13 with cumulative % formulas (0.35 pts)
  - Component 3: Column F has concatenation label formulas for all 12 data rows (0.30 pts)
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_formula_pattern_concat_005'


def normalize_formula(f):
    """Strip whitespace and uppercase for loose comparison."""
    if f is None:
        return ''
    return str(f).strip().upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition gate: file must load
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: 'Revenue' sheet must exist
    if 'Revenue' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Revenue' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Revenue']

    # -----------------------------------------------------------------
    # Component 1: Column D (Running Total) filled for rows 3-13 (0.35 pts)
    # The initial file only has D2 = '=SUM($B$2:B2)'. The task requires
    # filling D3:D13 with analogous formulas: =SUM($B$2:B3) ... =SUM($B$2:B13)
    # -----------------------------------------------------------------
    try:
        d_filled_count = 0
        d_total_rows = 11  # rows 3 through 13

        for row in range(3, 14):  # rows 3-13
            cell_val = ws.cell(row=row, column=4).value  # column D = 4
            if cell_val is None:
                print(f"FAIL: D{row} is empty — running total not filled down")
                continue
            norm = normalize_formula(cell_val)
            # Accept any formula that starts with =SUM($B$2:B{row} (with or without $)
            # Pattern: =SUM($B$2:B{row}) or =SUM($B$2:$B${row})
            expected_relative = f'=SUM($B$2:B{row})'
            expected_absolute = f'=SUM($B$2:$B${row})'
            if (norm == normalize_formula(expected_relative) or
                    norm == normalize_formula(expected_absolute) or
                    norm.startswith('=SUM($B$2:')):
                d_filled_count += 1
            else:
                print(f"FAIL: D{row} has unexpected formula: {cell_val!r}")

        if d_filled_count == d_total_rows:
            print(f"PASS: Component 1 — Column D running total filled for all {d_total_rows} rows (rows 3-13) (0.35 pts)")
            total_score += 0.35
        elif d_filled_count >= 1:
            partial = round(0.35 * d_filled_count / d_total_rows, 4)
            print(f"PARTIAL: Component 1 — Column D filled for {d_filled_count}/{d_total_rows} rows ({partial} pts)")
            total_score += partial
        else:
            print("FAIL: Component 1 — Column D not filled for any rows beyond row 2")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------
    # Component 2: Column E (Cumulative %) filled for rows 3-13 (0.35 pts)
    # The initial file only has E2 = '=D2/SUM($B$2:$B$13)'. The task
    # requires E3:E13 with analogous formulas referencing each row's D.
    # -----------------------------------------------------------------
    try:
        e_filled_count = 0
        e_total_rows = 11  # rows 3 through 13

        for row in range(3, 14):  # rows 3-13
            cell_val = ws.cell(row=row, column=5).value  # column E = 5
            if cell_val is None:
                print(f"FAIL: E{row} is empty — cumulative % not filled down")
                continue
            norm = normalize_formula(cell_val)
            # Accept any formula referencing D{row} / SUM($B$2:$B$13)
            # e.g., =D3/SUM($B$2:$B$13)
            expected = f'=D{row}/SUM($B$2:$B$13)'
            if norm == normalize_formula(expected) or (
                f'D{row}' in norm and 'SUM($B$2:$B$13)' in norm.upper()):
                e_filled_count += 1
            else:
                print(f"FAIL: E{row} has unexpected formula: {cell_val!r}")

        if e_filled_count == e_total_rows:
            print(f"PASS: Component 2 — Column E cumulative % filled for all {e_total_rows} rows (rows 3-13) (0.35 pts)")
            total_score += 0.35
        elif e_filled_count >= 1:
            partial = round(0.35 * e_filled_count / e_total_rows, 4)
            print(f"PARTIAL: Component 2 — Column E filled for {e_filled_count}/{e_total_rows} rows ({partial} pts)")
            total_score += partial
        else:
            print("FAIL: Component 2 — Column E not filled for any rows beyond row 2")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------
    # Component 3: Column F contains label concatenation formulas for
    # all 12 data rows (F2:F13) (0.30 pts)
    # Each F cell should contain a formula combining:
    #   "Month: " & A# & " | Rev: " & TEXT(B#,"0.00") &
    #   " | Total: " & TEXT(D#,"0.00") & " | Cum%: " & TEXT(E#,"0.00")
    # -----------------------------------------------------------------
    try:
        f_filled_count = 0
        f_total_rows = 12  # rows 2 through 13

        for row in range(2, 14):  # rows 2-13
            cell_val = ws.cell(row=row, column=6).value  # column F = 6
            if cell_val is None:
                print(f"FAIL: F{row} is empty — label concat not present")
                continue
            if not isinstance(cell_val, str) or not cell_val.startswith('='):
                print(f"FAIL: F{row} is not a formula: {cell_val!r}")
                continue
            norm = normalize_formula(cell_val)
            # Must reference A{row}, B{row}, D{row}, E{row} and use TEXT()
            # and contain "Month:" label pattern
            has_month_label = 'MONTH:' in norm or '"MONTH:"' in norm or '"MONTH: "' in norm
            has_rev_label = 'REV:' in norm or '"REV:"' in norm or '"REV: "' in norm
            has_total_label = 'TOTAL:' in norm or '"TOTAL:"' in norm or '"TOTAL: "' in norm
            has_cum_label = 'CUM%:' in norm or '"CUM%:"' in norm or '"CUM%: "' in norm
            has_text_func = 'TEXT(' in norm
            has_row_refs = (f'A{row}' in norm and f'B{row}' in norm and
                           f'D{row}' in norm and f'E{row}' in norm)
            if (has_month_label and has_rev_label and has_total_label and
                    has_cum_label and has_text_func and has_row_refs):
                f_filled_count += 1
            else:
                print(f"FAIL: F{row} formula missing expected components: {cell_val!r}")
                print(f"       has_month={has_month_label}, has_rev={has_rev_label}, "
                      f"has_total={has_total_label}, has_cum={has_cum_label}, "
                      f"has_text={has_text_func}, has_row_refs={has_row_refs}")

        if f_filled_count == f_total_rows:
            print(f"PASS: Component 3 — Column F concatenation labels present for all {f_total_rows} rows (0.30 pts)")
            total_score += 0.30
        elif f_filled_count >= 1:
            partial = round(0.30 * f_filled_count / f_total_rows, 4)
            print(f"PARTIAL: Component 3 — Column F labels present for {f_filled_count}/{f_total_rows} rows ({partial} pts)")
            total_score += partial
        else:
            print("FAIL: Component 3 — Column F has no valid concatenation formulas")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in given env
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
