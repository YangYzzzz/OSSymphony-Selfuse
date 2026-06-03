"""
Reward Script: Left VLOOKUP with INDEX/MATCH
Task ID: calc_fma_index_match_left_051
Domain: libreoffice_calc
Scoring:
  Component 1: B24:B33 are all non-empty (contain formulas/values) — 0.4 points
  Component 2: Formulas use INDEX/MATCH structure referencing correct ranges — 0.4 points
  Component 3: Formulas reference each row's name cell (A24, A25, ..., A33) — 0.2 points
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — reward scripts run on the VM
TASK_ID = 'calc_fma_index_match_left_051'


def verify_task(file_path):
    """
    Verify that cells B24:B33 in the 'Lookup' sheet contain INDEX/MATCH formulas
    that perform a 'left VLOOKUP' — looking up employee IDs (column A) by matching
    employee names (column B) in rows 2-21 of the lookup table.

    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the 'Lookup' sheet exists
    if 'Lookup' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Lookup' not found in workbook.")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Lookup']

    # Component 1: All 10 cells B24:B33 are non-empty (contain formulas) (0.4 points)
    # This FAILS on initial (all None) → PASSES on golden (all formulas)
    try:
        filled_count = 0
        empty_cells = []
        for row in range(24, 34):  # rows 24..33 inclusive
            cell_val = ws.cell(row=row, column=2).value
            if cell_val is not None and str(cell_val).strip() != '':
                filled_count += 1
            else:
                empty_cells.append(f"B{row}")

        if filled_count == 10:
            print(f"PASS: Component 1 — All 10 cells B24:B33 are populated (0.4 pts)")
            total_score += 0.4
        elif filled_count > 0:
            # Partial — some cells filled
            partial = round(0.4 * filled_count / 10, 2)
            print(f"PARTIAL: Component 1 — {filled_count}/10 cells populated, empty: {empty_cells} ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No cells in B24:B33 are populated (all empty)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Formulas use INDEX/MATCH structure referencing the correct lookup ranges (0.4 points)
    # Expected pattern: =INDEX($A$2:$A$21,MATCH(A##,$B$2:$B$21,0))
    # This FAILS on initial (no formulas) → PASSES on golden (formulas present)
    try:
        correct_structure_count = 0
        wrong_structure = []

        for row in range(24, 34):
            cell_val = ws.cell(row=row, column=2).value
            if cell_val is None:
                wrong_structure.append(f"B{row}: empty")
                continue

            formula = str(cell_val).strip().upper().replace(' ', '')

            # Check for INDEX function referencing column A lookup range
            has_index = 'INDEX(' in formula
            # Check for MATCH function referencing column B lookup range
            has_match = 'MATCH(' in formula
            # Check for reference to the employee ID range (A2:A21)
            has_a_range = 'A$2:$A$21' in formula or 'A2:A21' in formula.replace('$', '')
            # Check for reference to the employee name range (B2:B21)
            has_b_range = 'B$2:$B$21' in formula or 'B2:B21' in formula.replace('$', '')

            if has_index and has_match and has_a_range and has_b_range:
                correct_structure_count += 1
            else:
                details = f"INDEX={has_index}, MATCH={has_match}, A-range={has_a_range}, B-range={has_b_range}"
                wrong_structure.append(f"B{row}: {details} | formula={cell_val}")

        if correct_structure_count == 10:
            print(f"PASS: Component 2 — All 10 formulas use INDEX/MATCH with correct lookup ranges (0.4 pts)")
            total_score += 0.4
        elif correct_structure_count > 0:
            partial = round(0.4 * correct_structure_count / 10, 2)
            print(f"PARTIAL: Component 2 — {correct_structure_count}/10 formulas have correct structure ({partial} pts)")
            for ws_detail in wrong_structure:
                print(f"  Wrong: {ws_detail}")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No formulas have the correct INDEX/MATCH structure")
            for ws_detail in wrong_structure[:3]:
                print(f"  {ws_detail}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Each formula references the correct row's name cell (A24, A25, ..., A33) (0.2 points)
    # This ensures each formula is row-specific, not a copy-paste error referencing the wrong row.
    # This FAILS on initial (no formulas) → PASSES on golden (each formula references its own row)
    try:
        row_specific_count = 0
        row_errors = []

        for row in range(24, 34):
            cell_val = ws.cell(row=row, column=2).value
            if cell_val is None:
                row_errors.append(f"B{row}: empty")
                continue

            formula = str(cell_val).strip()
            # The MATCH lookup value should reference the name cell in the same row
            # e.g., row 24 should reference A24, row 25 should reference A25, etc.
            # Pattern: MATCH(A{row}, ...) — case insensitive
            formula_upper = formula.upper()
            expected_ref = f'A{row}'

            # Check that the formula contains the row-specific reference
            # Allow for absolute reference like A$24 or $A$24 as well
            has_row_ref = (
                f'A{row},' in formula_upper or
                f'A{row})' in formula_upper or
                f'A${row},' in formula_upper or
                f'A${row})' in formula_upper or
                f'$A{row},' in formula_upper or
                f'$A{row})' in formula_upper or
                f'$A${row},' in formula_upper or
                f'$A${row})' in formula_upper
            )

            if has_row_ref:
                row_specific_count += 1
            else:
                row_errors.append(f"B{row}: expected reference to A{row}, formula={formula}")

        if row_specific_count == 10:
            print(f"PASS: Component 3 — All 10 formulas reference their row-specific name cell (0.2 pts)")
            total_score += 0.2
        elif row_specific_count > 0:
            partial = round(0.2 * row_specific_count / 10, 2)
            print(f"PARTIAL: Component 3 — {row_specific_count}/10 formulas reference correct row cell ({partial} pts)")
            for err in row_errors[:3]:
                print(f"  {err}")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No formulas reference row-specific name cells")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
