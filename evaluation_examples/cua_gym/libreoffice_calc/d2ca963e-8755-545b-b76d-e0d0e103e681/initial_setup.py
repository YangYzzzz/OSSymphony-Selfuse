"""
Initial Setup: INDEX/MATCH left lookup task
Task ID: calc_fma_index_match_left_051
Domain: libreoffice_calc

Creates a spreadsheet with a lookup table (Employee IDs and Names)
and a search area where the user needs to enter INDEX/MATCH formulas.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fma_index_match_left_051'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Lookup'

    # --- Header row for lookup table ---
    ws['A1'] = 'Employee ID'
    ws['B1'] = 'Employee Name'
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)

    # --- Lookup table: rows 2-21 ---
    # 20 unique employees: EMP001-EMP020
    employees = [
        ('EMP001', 'Alice Brown'),
        ('EMP002', 'Bob Smith'),
        ('EMP003', 'Carol Jones'),
        ('EMP004', 'David Lee'),
        ('EMP005', 'Eve Davis'),
        ('EMP006', 'Frank Miller'),
        ('EMP007', 'Grace Wilson'),
        ('EMP008', 'Hank Moore'),
        ('EMP009', 'Iris Taylor'),
        ('EMP010', 'Jack Anderson'),
        ('EMP011', 'Karen Thomas'),
        ('EMP012', 'Leo Jackson'),
        ('EMP013', 'Mia White'),
        ('EMP014', 'Nathan Harris'),
        ('EMP015', 'Olivia Martin'),
        ('EMP016', 'Peter Thompson'),
        ('EMP017', 'Quinn Garcia'),
        ('EMP018', 'Rachel Martinez'),
        ('EMP019', 'Samuel Robinson'),
        ('EMP020', 'Tina Clark'),
    ]

    for i, (emp_id, emp_name) in enumerate(employees, 2):
        ws.cell(row=i, column=1, value=emp_id)
        ws.cell(row=i, column=2, value=emp_name)

    # --- Row 22: blank separator ---

    # --- Row 23: Search area header ---
    ws['A23'] = 'Name'
    ws['B23'] = 'Employee ID'
    ws['A23'].font = Font(bold=True)
    ws['B23'].font = Font(bold=True)

    # --- Rows 24-33: names to look up (Column A), empty Column B ---
    lookup_names = [
        'Alice Brown',
        'Bob Smith',
        'Carol Jones',
        'David Lee',
        'Eve Davis',
        'Frank Miller',
        'Grace Wilson',
        'Hank Moore',
        'Iris Taylor',
        'Jack Anderson',
    ]

    for i, name in enumerate(lookup_names, 24):
        ws.cell(row=i, column=1, value=name)
        # Column B intentionally left empty for user to fill with INDEX/MATCH

    # --- Column widths ---
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Lookup table: rows 2-21 (EMP001-EMP020 with employee names)')
    print(f'  Search area: rows 24-33 (names in col A, col B empty)')


create_initial()
