"""
Initial Setup: Create multi-sheet regional sales workbook with no named ranges
Task ID: calc_nrv_039
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_039'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

MONTHS = [
    'January', 'February', 'March', 'April', 'May', 'June',
    'July', 'August', 'September', 'October', 'November', 'December'
]

# Realistic monthly sales data for four regions
REGION_DATA = {
    'North': [45230, 38920, 52100, 49870, 61340, 58790,
              63210, 67450, 55890, 48760, 71230, 82150],
    'South': [32150, 29870, 35640, 41230, 38970, 44560,
              47890, 51230, 43670, 39450, 52340, 61780],
    'East':  [51890, 48320, 55670, 53210, 62450, 59830,
              64780, 69120, 57340, 52190, 73450, 85670],
    'West':  [28970, 25640, 31450, 35780, 33210, 39450,
              42670, 46890, 38120, 34560, 48970, 57230],
}


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # Create regional sheets
    first = True
    for region, sales in REGION_DATA.items():
        if first:
            ws = wb.active
            ws.title = region
            first = False
        else:
            ws = wb.create_sheet(region)

        ws.cell(row=1, column=1, value='Month')
        ws.cell(row=1, column=2, value='Sales')

        for i, (month, amount) in enumerate(zip(MONTHS, sales), start=2):
            ws.cell(row=i, column=1, value=month)
            ws.cell(row=i, column=2, value=amount)

        # Set reasonable column widths
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 12

    # Create Consolidated sheet
    ws_con = wb.create_sheet('Consolidated')
    ws_con.cell(row=1, column=1, value='Region Total')
    ws_con.cell(row=1, column=2, value='Amount')
    ws_con.cell(row=2, column=1, value='All Regions')
    # B2 intentionally left empty — the task is to fill it
    ws_con.column_dimensions['A'].width = 16
    ws_con.column_dimensions['B'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
