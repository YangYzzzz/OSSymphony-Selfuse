"""
Reward Script: Insert 'Region Code' column before B, label B1, apply double-line right border
Task ID: calc_gg1_046
Domain: libreoffice_calc
Scoring:
  Component 1: B1 header is 'Region Code' (0.25)
  Component 2: Original data shifted right - C1 is 'Salesperson' (0.25)
  Component 3: Column count increased to 6 (0.10)
  Component 4: Double-line right border on column B cells (0.40)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_046'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Sales' sheet must exist
    if 'Sales' not in wb.sheetnames:
        print("FAIL: 'Sales' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Sales']

    # Component 1: B1 header is 'Region Code' (0.25 points)
    # This verifies a new column was inserted and labeled correctly.
    # In initial_env, B1 = 'Salesperson', so this FAILS on initial.
    try:
        b1_val = ws['B1'].value
        if b1_val is not None and str(b1_val).strip() == 'Region Code':
            print(f"PASS: Component 1 — B1 is 'Region Code' (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — B1 expected 'Region Code', found: {b1_val!r}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Original data shifted right — C1 is 'Salesperson' (0.25 points)
    # In initial_env, C1 = 'Product' (not 'Salesperson'), so this FAILS on initial.
    try:
        c1_val = ws['C1'].value
        if c1_val is not None and str(c1_val).strip() == 'Salesperson':
            print(f"PASS: Component 2 — C1 is 'Salesperson' (data shifted correctly) (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — C1 expected 'Salesperson', found: {c1_val!r}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Column count is 6 (was 5 in initial) (0.10 points)
    # In initial_env, max_column = 5, so checking for 6 FAILS on initial.
    try:
        max_col = ws.max_column
        if max_col >= 6:
            print(f"PASS: Component 3 — Column count is {max_col} (>= 6) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — Expected >= 6 columns, found: {max_col}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Double-line right border on column B cells (0.40 points)
    # Check a sample of cells in column B for double right border.
    # In initial_env, column B has no special borders, so this FAILS on initial.
    try:
        # Sample rows to check: header + several data rows + deep rows
        sample_rows = [1, 2, 5, 10, 16]
        # Also add some deeper rows if they exist
        if ws.max_row >= 100:
            sample_rows.extend([50, 100])
        if ws.max_row >= 500:
            sample_rows.append(500)

        double_count = 0
        checked = 0
        for r in sample_rows:
            if r <= ws.max_row:
                cell = ws.cell(row=r, column=2)
                border_style = cell.border.right.style
                checked += 1
                if border_style == 'double':
                    double_count += 1
                else:
                    print(f"  B{r} border right style: {border_style!r} (expected 'double')")

        if checked > 0:
            ratio = double_count / checked
            if ratio >= 0.8:
                # At least 80% of sampled cells have double right border
                print(f"PASS: Component 4 — Double right border on {double_count}/{checked} sampled B cells (0.40 pts)")
                total_score += 0.40
            elif ratio >= 0.5:
                # Partial credit for partial application
                partial = round(0.40 * ratio, 2)
                print(f"PARTIAL: Component 4 — Double right border on {double_count}/{checked} sampled B cells ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 4 — Double right border on only {double_count}/{checked} sampled B cells")
        else:
            print("FAIL: Component 4 — No cells to check in column B")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
