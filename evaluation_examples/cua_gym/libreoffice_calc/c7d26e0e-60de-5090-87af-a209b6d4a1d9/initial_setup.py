"""
Initial Setup: Insert Region Code column with double-line border
Task ID: calc_gg1_046
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_046'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    # Headers: A=Sale ID, B=Salesperson, C=Product, D=Amount, E=Date
    headers = ['Sale ID', 'Salesperson', 'Product', 'Amount', 'Date']
    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Realistic sales data (15 rows)
    data = [
        ['S-1001', 'Sarah Chen', 'Enterprise License', 45230.00, '2025-03-15'],
        ['S-1002', 'Marcus Johnson', 'Cloud Hosting Plan', 12750.00, '2025-03-16'],
        ['S-1003', 'Elena Rodriguez', 'Data Analytics Suite', 28900.00, '2025-03-17'],
        ['S-1004', 'James Okafor', 'Security Package', 15400.00, '2025-03-18'],
        ['S-1005', 'Priya Patel', 'Enterprise License', 45230.00, '2025-03-19'],
        ['S-1006', 'David Kim', 'Cloud Hosting Plan', 18500.00, '2025-03-20'],
        ['S-1007', 'Rachel Thompson', 'Support Contract', 8200.00, '2025-03-21'],
        ['S-1008', 'Ahmed Hassan', 'Data Analytics Suite', 32100.00, '2025-03-22'],
        ['S-1009', 'Lisa Nakamura', 'Security Package', 15400.00, '2025-03-23'],
        ['S-1010', 'Carlos Mendez', 'Enterprise License', 52000.00, '2025-03-24'],
        ['S-1011', 'Sophie Martin', 'Cloud Hosting Plan', 14300.00, '2025-03-25'],
        ['S-1012', 'Wei Zhang', 'Support Contract', 9750.00, '2025-03-26'],
        ['S-1013', 'Olivia Brown', 'Data Analytics Suite', 27600.00, '2025-03-27'],
        ['S-1014', 'Nathan Williams', 'Security Package', 19800.00, '2025-03-28'],
        ['S-1015', 'Fatima Al-Rashid', 'Enterprise License', 48500.00, '2025-03-29'],
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])  # Sale ID
        ws.cell(row=r, column=2, value=row_data[1])  # Salesperson
        ws.cell(row=r, column=3, value=row_data[2])  # Product
        cell_amount = ws.cell(row=r, column=4, value=row_data[3])  # Amount
        cell_amount.number_format = '$#,##0.00'
        ws.cell(row=r, column=5, value=row_data[4])  # Date

    # Set column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
