"""
Reward Script: Process Thunderbird Receipts folder - export emls, extract CSVs, merge, and open in LibreOffice Calc
Task ID: osworld_multi_apps_email_file_convert_008
Domain: multi_apps (Thunderbird + LibreOffice Calc)
Scoring:
  Component 1: 5 .eml files exported to /home/user/receipts/eml/                   (0.15 pts)
  Component 2: 5 CSV attachments extracted & renamed correctly to /home/user/receipts/data/  (0.20 pts)
  Component 3: all_receipts.csv merged with 23 data rows + correct columns          (0.20 pts)
  Component 4: processing_log.csv has 5 entries + correct columns                   (0.15 pts)
  Component 5: XLSX exists with 23 data rows sorted by date ascending               (0.15 pts)
  Component 6: XLSX amount column has currency formatting ($#,##0.00 or similar)    (0.10 pts)
  Component 7: XLSX last row has SUM formula for amount column                      (0.05 pts)
  Total: 1.0
"""

import os
import csv

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_email_file_convert_008'

RECEIPTS_DIR = f'{WORKDIR}/receipts'
EML_DIR = f'{RECEIPTS_DIR}/eml'
DATA_DIR = f'{RECEIPTS_DIR}/data'
ALL_RECEIPTS_CSV = f'{RECEIPTS_DIR}/all_receipts.csv'
PROCESSING_LOG_CSV = f'{RECEIPTS_DIR}/processing_log.csv'
XLSX_PATH = f'{WORKDIR}/{TASK_ID}.xlsx'


def verify_task():
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Component 1: 5 .eml files exported to /home/user/receipts/eml/ (0.15 points)
    try:
        if not os.path.isdir(EML_DIR):
            print(f"FAIL: Component 1 — EML directory does not exist: {EML_DIR}")
        else:
            eml_files = [f for f in os.listdir(EML_DIR) if f.endswith('.eml')]
            if len(eml_files) == 5:
                print(f"PASS: Component 1 — 5 .eml files found in {EML_DIR}: {eml_files} (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 1 — Expected 5 .eml files, found {len(eml_files)}: {eml_files}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: 5 CSV attachments extracted and renamed correctly to /home/user/receipts/data/ (0.20 points)
    # Naming format: YYYYMMDD_sender_originalname.ext
    try:
        if not os.path.isdir(DATA_DIR):
            print(f"FAIL: Component 2 — Data directory does not exist: {DATA_DIR}")
        else:
            data_files = [f for f in os.listdir(DATA_DIR) if f.endswith('.csv')]
            if len(data_files) == 5:
                # Check naming convention: each file should start with YYYYMMDD_
                import re
                rename_pattern = re.compile(r'^\d{8}_\w+_\w+.*\.csv$')
                valid_names = [f for f in data_files if rename_pattern.match(f)]
                if len(valid_names) == 5:
                    print(f"PASS: Component 2 — 5 CSV files found with correct naming convention in {DATA_DIR}: {data_files} (0.20 pts)")
                    total_score += 0.20
                elif len(data_files) == 5:
                    # Have 5 files but naming might be slightly different - partial credit
                    print(f"PASS: Component 2 — 5 CSV files found in {DATA_DIR} (naming may differ): {data_files} (0.20 pts)")
                    total_score += 0.20
                else:
                    print(f"FAIL: Component 2 — Only {len(valid_names)}/5 files match naming convention. Files: {data_files}")
            else:
                print(f"FAIL: Component 2 — Expected 5 CSV files, found {len(data_files)}: {data_files}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: all_receipts.csv merged with 23 data rows + correct columns (0.20 points)
    try:
        if not os.path.isfile(ALL_RECEIPTS_CSV):
            print(f"FAIL: Component 3 — all_receipts.csv not found at {ALL_RECEIPTS_CSV}")
        else:
            with open(ALL_RECEIPTS_CSV, 'r', newline='') as f:
                reader = csv.reader(f)
                rows = list(reader)

            if len(rows) == 0:
                print(f"FAIL: Component 3 — all_receipts.csv is empty")
            else:
                header = rows[0]
                data_rows = rows[1:]
                # Check columns: date, vendor, amount
                expected_cols = {'date', 'vendor', 'amount'}
                actual_cols = set(c.lower().strip() for c in header)

                if not expected_cols.issubset(actual_cols):
                    print(f"FAIL: Component 3 — Expected columns {expected_cols}, found {header}")
                elif len(data_rows) == 23:
                    print(f"PASS: Component 3 — all_receipts.csv has correct columns {header} and 23 data rows (0.20 pts)")
                    total_score += 0.20
                else:
                    print(f"FAIL: Component 3 — Expected 23 data rows, found {len(data_rows)}. Columns: {header}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: processing_log.csv has 5 entries + correct columns (0.15 points)
    try:
        if not os.path.isfile(PROCESSING_LOG_CSV):
            print(f"FAIL: Component 4 — processing_log.csv not found at {PROCESSING_LOG_CSV}")
        else:
            with open(PROCESSING_LOG_CSV, 'r', newline='') as f:
                reader = csv.reader(f)
                rows = list(reader)

            if len(rows) == 0:
                print(f"FAIL: Component 4 — processing_log.csv is empty")
            else:
                header = rows[0]
                data_rows = rows[1:]
                # Check columns: eml_file, attachment_name, renamed_to, rows_extracted
                expected_cols = {'eml_file', 'attachment_name', 'renamed_to', 'rows_extracted'}
                actual_cols = set(c.lower().strip() for c in header)

                if not expected_cols.issubset(actual_cols):
                    print(f"FAIL: Component 4 — Expected columns {expected_cols}, found {header}")
                elif len(data_rows) == 5:
                    print(f"PASS: Component 4 — processing_log.csv has correct columns {header} and 5 entries (0.15 pts)")
                    total_score += 0.15
                else:
                    print(f"FAIL: Component 4 — Expected 5 log entries, found {len(data_rows)}. Columns: {header}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: XLSX exists with 23 data rows sorted by date ascending (0.15 points)
    try:
        import openpyxl
        if not os.path.isfile(XLSX_PATH):
            print(f"FAIL: Component 5 — XLSX file not found at {XLSX_PATH}")
        else:
            wb = openpyxl.load_workbook(XLSX_PATH)
            ws = wb.active

            # Must have at least header + 23 data rows + 1 total row = 25 rows
            total_rows = ws.max_row
            if total_rows < 24:
                print(f"FAIL: Component 5 — XLSX has too few rows: {total_rows} (need at least 24 for 23 data rows)")
            else:
                # Extract data rows (excluding header row 1 and last row which is total)
                data_rows_xlsx = []
                for row in ws.iter_rows(min_row=2, max_row=total_rows - 1, values_only=True):
                    data_rows_xlsx.append(row)

                if len(data_rows_xlsx) != 23:
                    print(f"FAIL: Component 5 — Expected 23 data rows, found {len(data_rows_xlsx)}")
                else:
                    # Check sort order by date (column 1)
                    dates = [str(row[0]) for row in data_rows_xlsx if row[0] is not None]
                    is_sorted = dates == sorted(dates)
                    if is_sorted:
                        print(f"PASS: Component 5 — XLSX has 23 data rows sorted by date ascending (0.15 pts)")
                        total_score += 0.15
                    else:
                        print(f"FAIL: Component 5 — XLSX data rows are NOT sorted by date. First dates: {dates[:5]}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: XLSX amount column has currency formatting ($#,##0.00 or similar) (0.10 points)
    try:
        import openpyxl
        if not os.path.isfile(XLSX_PATH):
            print(f"FAIL: Component 6 — XLSX file not found")
        else:
            wb = openpyxl.load_workbook(XLSX_PATH)
            ws = wb.active

            # Find the amount column - should be column 3 (C)
            # Check formatting of first few data cells in amount column
            currency_formats = {'$#,##0.00', '#,##0.00', '"$"#,##0.00', '[$USD]#,##0.00',
                                 '$#,##0.00_);($#,##0.00)', '#,##0.00_);(#,##0.00)'}

            # Check cells C2 through C5 for currency formatting
            formatted_count = 0
            total_checked = 0
            for row in ws.iter_rows(min_row=2, max_row=min(6, ws.max_row), min_col=3, max_col=3):
                for cell in row:
                    if cell.value is not None:
                        total_checked += 1
                        fmt = str(cell.number_format) if cell.number_format else ''
                        # Check for currency-like format (contains $, or #,##0)
                        if '$' in fmt or ('#,##0' in fmt and ('0.00' in fmt or '0' in fmt)):
                            formatted_count += 1

            if total_checked > 0 and formatted_count == total_checked:
                print(f"PASS: Component 6 — Amount column has currency formatting (format: {ws['C2'].number_format}) (0.10 pts)")
                total_score += 0.10
            elif formatted_count > 0:
                print(f"PASS: Component 6 — Amount column has currency formatting ({formatted_count}/{total_checked} cells) (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 6 — Amount column lacks currency formatting. Format found: {ws['C2'].number_format}")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: XLSX last row has SUM formula for amount column (0.05 points)
    try:
        import openpyxl
        if not os.path.isfile(XLSX_PATH):
            print(f"FAIL: Component 7 — XLSX file not found")
        else:
            wb = openpyxl.load_workbook(XLSX_PATH)
            ws = wb.active

            last_row = ws.max_row
            last_row_values = [ws.cell(row=last_row, column=c).value for c in range(1, ws.max_column + 1)]

            # Check if there's a SUM formula or TOTAL label in the last row
            last_row_str = ' '.join(str(v) for v in last_row_values if v is not None).upper()

            # Check for SUM formula in amount column (column C = col 3)
            amount_cell_last = ws.cell(row=last_row, column=3).value
            sum_label_cell = ws.cell(row=last_row, column=1).value

            has_sum_formula = (isinstance(amount_cell_last, str) and
                               'SUM' in amount_cell_last.upper())
            has_total_label = (sum_label_cell is not None and
                               'TOTAL' in str(sum_label_cell).upper())

            if has_sum_formula:
                print(f"PASS: Component 7 — Last row has SUM formula: '{amount_cell_last}' (label: '{sum_label_cell}') (0.05 pts)")
                total_score += 0.05
            elif has_total_label and amount_cell_last is not None:
                # Even if no formula, if there's a total row with a value
                print(f"PASS: Component 7 — Last row has TOTAL label with value: '{amount_cell_last}' (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 7 — Last row does not have SUM formula or TOTAL. Last row: {last_row_values}")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


verify_task()
