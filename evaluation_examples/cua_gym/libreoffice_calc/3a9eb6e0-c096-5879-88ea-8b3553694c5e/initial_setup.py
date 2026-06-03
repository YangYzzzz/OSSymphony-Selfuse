"""
Initial Setup: Format data table with outer and inner borders
Task ID: calc_fmt_border_inner_outer_distinct_097
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_border_inner_outer_distinct_097'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Summary Table ---
    ws = wb.active
    ws.title = 'Summary Table'

    # Headers in row 1
    headers = ['Category', 'Q1', 'Q2', 'Q3', 'Q4']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data rows 2-12 (11 rows of realistic business data)
    data = [
        ['Electronics',    142500, 158300, 175200, 192400],
        ['Apparel',         87300,  92100,  88700,  96500],
        ['Home & Garden',   54200,  48900,  63100,  71800],
        ['Sports & Outdoor', 39800, 44600,  52300,  61200],
        ['Books & Media',   28400,  31200,  29800,  33600],
        ['Toys & Games',    46700,  51300,  48900,  87400],
        ['Health & Beauty', 62100,  67400,  71200,  75900],
        ['Automotive',      33500,  29800,  31400,  28700],
        ['Food & Beverage', 118200, 124600, 131800, 143200],
        ['Office Supplies',  21400,  23100,  24800,  27300],
        ['Jewelry',          15600,  18200,  21900,  26400],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name='Calibri', size=11)
            if c > 1:
                cell.number_format = '#,##0'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12

    # Row 1 height
    ws.row_dimensions[1].height = 18

    # NOTE: No borders applied — task requires adding them

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
