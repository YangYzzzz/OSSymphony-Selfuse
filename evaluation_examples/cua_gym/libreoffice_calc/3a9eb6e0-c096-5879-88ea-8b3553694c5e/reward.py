"""
Reward Script: Format data table A1:E12 with thick outer border and thin inner borders
Task ID: calc_fmt_border_inner_outer_distinct_097
Domain: libreoffice_calc
Scoring:
  Component 1: Outer border - top edge (row 1 top side thick black)          0.25 pts
  Component 2: Outer border - remaining 3 edges (bot/left/right thick black) 0.25 pts
  Component 3: Inner borders (all interior sides thin gray #808080)           0.30 pts
  Component 4: Inner borders applied AND no spillover outside A1:E12          0.10 pts
  Component 5: Cell values preserved (anchored to borders being applied)      0.10 pts
Total: 1.0
"""

import os
import openpyxl
from openpyxl.utils import get_column_letter

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmt_border_inner_outer_distinct_097'

OUTER_STYLE = 'thick'
OUTER_COLOR = 'FF000000'
INNER_STYLE = 'thin'
INNER_COLOR = 'FF808080'

# Range boundaries
MIN_ROW, MAX_ROW = 1, 12
MIN_COL, MAX_COL = 1, 5  # A=1, E=5


def check_border_side(side, expected_style, expected_color=None):
    """Return True if a border side matches the expected style (and optionally color)."""
    if side.style != expected_style:
        return False
    if expected_color is not None:
        try:
            if side.color.rgb != expected_color:
                return False
        except Exception:
            return False
    return True


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet must exist
    if 'Summary Table' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Summary Table' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Summary Table']

    # ------------------------------------------------------------------
    # Component 1: Outer border — top edge (row 1, all columns A-E)
    # All cells in row 1 must have thick black top border
    # This FAILS on initial (no borders) → PASSES on golden
    # ------------------------------------------------------------------
    try:
        top_edge_failures = []
        for col in range(MIN_COL, MAX_COL + 1):
            cell = ws.cell(row=MIN_ROW, column=col)
            if not check_border_side(cell.border.top, OUTER_STYLE, OUTER_COLOR):
                top_edge_failures.append(f"{get_column_letter(col)}{MIN_ROW}.top={cell.border.top.style}")

        if len(top_edge_failures) == 0:
            print(f"PASS: Component 1 — Top outer border (row 1 top) is thick black on all 5 columns (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — Top outer border fails: {top_edge_failures[:3]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ------------------------------------------------------------------
    # Component 2: Outer border — bottom, left, right edges
    # Bottom of row 12, left of col A, right of col E — all thick black
    # This FAILS on initial → PASSES on golden
    # ------------------------------------------------------------------
    try:
        outer_failures = []

        # Bottom edge: row 12, all columns
        for col in range(MIN_COL, MAX_COL + 1):
            cell = ws.cell(row=MAX_ROW, column=col)
            if not check_border_side(cell.border.bottom, OUTER_STYLE, OUTER_COLOR):
                outer_failures.append(f"{get_column_letter(col)}{MAX_ROW}.bot={cell.border.bottom.style}")

        # Left edge: col A, all rows
        for row in range(MIN_ROW, MAX_ROW + 1):
            cell = ws.cell(row=row, column=MIN_COL)
            if not check_border_side(cell.border.left, OUTER_STYLE, OUTER_COLOR):
                outer_failures.append(f"A{row}.left={cell.border.left.style}")

        # Right edge: col E, all rows
        for row in range(MIN_ROW, MAX_ROW + 1):
            cell = ws.cell(row=row, column=MAX_COL)
            if not check_border_side(cell.border.right, OUTER_STYLE, OUTER_COLOR):
                outer_failures.append(f"E{row}.right={cell.border.right.style}")

        if not outer_failures:
            print(f"PASS: Component 2 — Bottom/left/right outer borders are thick black (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — Outer border failures: {outer_failures[:5]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ------------------------------------------------------------------
    # Component 3: Inner borders — all interior cell sides are thin gray
    # Interior sides = borders shared between two cells within the range
    # Specifically:
    #   - Horizontal interior: bottom of rows 1-11 (for cols A-E)
    #                          top of rows 2-12 (for cols A-E)
    #   - Vertical interior: right of cols A-D (for rows 1-12)
    #                        left of cols B-E (for rows 1-12)
    # But we exclude the outer border sides (already checked above).
    # Check interior-only sides:
    #   - For cells in col B-D (non-edge columns): all 4 sides = thin gray
    #   - For cells in col A (non-bottom-edge rows): right side = thin gray
    #   - For cells in col E (non-bottom-edge rows): left side = thin gray
    #   - Interior horizontal: bottom of rows 1-11 for any col
    # This FAILS on initial → PASSES on golden
    # ------------------------------------------------------------------
    try:
        inner_failures = []

        # Check all interior cells (B2:D11) — all 4 sides should be thin gray
        for row in range(MIN_ROW + 1, MAX_ROW):  # rows 2-11
            for col in range(MIN_COL + 1, MAX_COL):  # cols B-D (2-4)
                cell = ws.cell(row=row, column=col)
                b = cell.border
                for side_name, side in [('top', b.top), ('bottom', b.bottom),
                                        ('left', b.left), ('right', b.right)]:
                    if not check_border_side(side, INNER_STYLE, INNER_COLOR):
                        inner_failures.append(
                            f"{get_column_letter(col)}{row}.{side_name}={side.style}"
                        )

        # Check row 1 interior cells (B1:D1): bottom + left + right = thin gray
        # (top is already checked in Component 1 as outer)
        for col in range(MIN_COL + 1, MAX_COL):  # cols B-D
            cell = ws.cell(row=MIN_ROW, column=col)
            b = cell.border
            for side_name, side in [('bottom', b.bottom), ('left', b.left), ('right', b.right)]:
                if not check_border_side(side, INNER_STYLE, INNER_COLOR):
                    inner_failures.append(
                        f"{get_column_letter(col)}1.{side_name}={side.style}"
                    )

        # Check row 12 interior cells (B12:D12): top + left + right = thin gray
        # (bottom is outer)
        for col in range(MIN_COL + 1, MAX_COL):  # cols B-D
            cell = ws.cell(row=MAX_ROW, column=col)
            b = cell.border
            for side_name, side in [('top', b.top), ('left', b.left), ('right', b.right)]:
                if not check_border_side(side, INNER_STYLE, INNER_COLOR):
                    inner_failures.append(
                        f"{get_column_letter(col)}{MAX_ROW}.{side_name}={side.style}"
                    )

        # Check col A interior horizontal sides (A1:A11 bottom, A2:A12 top)
        # and col A right side (all rows) should be thin gray
        for row in range(MIN_ROW, MAX_ROW + 1):
            cell = ws.cell(row=row, column=MIN_COL)
            b = cell.border
            # right side = inner
            if not check_border_side(b.right, INNER_STYLE, INNER_COLOR):
                inner_failures.append(f"A{row}.right={b.right.style}")
            # bottom of rows 1-11 = inner horizontal
            if row < MAX_ROW:
                if not check_border_side(b.bottom, INNER_STYLE, INNER_COLOR):
                    inner_failures.append(f"A{row}.bottom={b.bottom.style}")
            # top of rows 2-12 = inner horizontal
            if row > MIN_ROW:
                if not check_border_side(b.top, INNER_STYLE, INNER_COLOR):
                    inner_failures.append(f"A{row}.top={b.top.style}")

        # Check col E right=outer (already done), left side = inner gray
        for row in range(MIN_ROW, MAX_ROW + 1):
            cell = ws.cell(row=row, column=MAX_COL)
            b = cell.border
            if not check_border_side(b.left, INNER_STYLE, INNER_COLOR):
                inner_failures.append(f"E{row}.left={b.left.style}")
            # bottom of rows 1-11 = inner horizontal
            if row < MAX_ROW:
                if not check_border_side(b.bottom, INNER_STYLE, INNER_COLOR):
                    inner_failures.append(f"E{row}.bottom={b.bottom.style}")
            if row > MIN_ROW:
                if not check_border_side(b.top, INNER_STYLE, INNER_COLOR):
                    inner_failures.append(f"E{row}.top={b.top.style}")

        if not inner_failures:
            print(f"PASS: Component 3 — All inner borders are thin gray (FF808080) (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 3 — Inner border failures ({len(inner_failures)} issues): {inner_failures[:5]}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ------------------------------------------------------------------
    # Component 4: No borders outside A1:E12 AND at least one inner border present
    # This is a compound check anchored to the task change:
    # - The table A1:E12 must have borders applied (inner gray borders exist)
    # - AND no spillover borders exist outside the range
    # Since the initial file has NO borders anywhere (inside or outside),
    # we anchor to inner borders being applied first. The "no spillover" check
    # only awards points if borders have been applied within the range.
    # On initial: inner borders are absent, so no points awarded here.
    # On golden: inner borders present AND no spillover.
    # ------------------------------------------------------------------
    try:
        # First, confirm at least one inner border is present (anchors to task completion)
        sample_inner_cell = ws.cell(row=2, column=2)
        inner_border_applied = (
            sample_inner_cell.border.top.style == INNER_STYLE and
            sample_inner_cell.border.left.style == INNER_STYLE
        )

        if not inner_border_applied:
            print(f"FAIL: Component 4 — No inner borders detected; cannot verify containment")
        else:
            # Inner borders exist; now check no spillover outside range
            outside_failures = []
            for col in range(MIN_COL, MAX_COL + 2):
                cell = ws.cell(row=MAX_ROW + 1, column=col)
                b = cell.border
                for side_name, side in [('top', b.top), ('bottom', b.bottom),
                                         ('left', b.left), ('right', b.right)]:
                    if side.style is not None:
                        outside_failures.append(f"{get_column_letter(col)}{MAX_ROW+1}.{side_name}={side.style}")

            for row in range(MIN_ROW, MAX_ROW + 1):
                cell = ws.cell(row=row, column=MAX_COL + 1)
                b = cell.border
                for side_name, side in [('top', b.top), ('bottom', b.bottom),
                                         ('left', b.left), ('right', b.right)]:
                    if side.style is not None:
                        outside_failures.append(f"F{row}.{side_name}={side.style}")

            if not outside_failures:
                print(f"PASS: Component 4 — Inner borders applied and no spillover outside A1:E12 (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 4 — Borders found outside range: {outside_failures[:3]}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ------------------------------------------------------------------
    # Component 5: Cell values preserved — verified jointly with outer border
    # Only awards points when the outer AND inner borders are both applied,
    # ensuring we're measuring a post-task state, not the initial state.
    # On initial: outer borders are absent → borders_fully_applied is False → 0 pts
    # On golden: borders applied AND values intact → 0.10 pts
    # ------------------------------------------------------------------
    try:
        expected_headers = ['Category', 'Q1', 'Q2', 'Q3', 'Q4']
        header_failures = []
        for col_idx, expected in enumerate(expected_headers, 1):
            actual = ws.cell(row=1, column=col_idx).value
            if str(actual).strip() != expected:
                header_failures.append(
                    f"{get_column_letter(col_idx)}1={repr(actual)}, expected={repr(expected)}"
                )

        # Outer border must be applied (thick black on top of A1) AND inner borders applied
        # This anchors to task completion, ensuring initial file doesn't score here
        outer_applied = check_border_side(ws.cell(row=1, column=1).border.top, OUTER_STYLE, OUTER_COLOR)
        inner_applied = check_border_side(ws.cell(row=2, column=2).border.top, INNER_STYLE, INNER_COLOR)
        borders_fully_applied = outer_applied and inner_applied

        if len(header_failures) == 0 and borders_fully_applied:
            print(f"PASS: Component 5 — Cell values preserved (headers intact, borders confirmed applied) (0.10 pts)")
            total_score += 0.10
        elif not borders_fully_applied:
            print(f"FAIL: Component 5 — Borders not fully applied; skipping value check")
        else:
            print(f"FAIL: Component 5 — Header values changed: {header_failures}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
