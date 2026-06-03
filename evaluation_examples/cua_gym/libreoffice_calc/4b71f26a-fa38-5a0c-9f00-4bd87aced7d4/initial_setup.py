"""
Initial Setup: Conditional formatting applied to entire column A instead of A2:A100
Task ID: calc_tbl_026
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_026'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Employees ---
    ws = wb.active
    ws.title = "Employees"

    # Headers
    headers = ["Name", "Department", "Salary", "Start Date", "Performance Rating"]
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Realistic employee data (99 rows, A2:A100)
    departments = ["Engineering", "Marketing", "Sales", "Finance", "HR", "Operations", "Legal", "Support"]
    ratings = ["Exceeds Expectations", "Meets Expectations", "Needs Improvement", "Outstanding", "Satisfactory"]

    employees = [
        ("Sarah Chen", "Engineering", 95000, "2021-03-15", "Outstanding"),
        ("Marcus Johnson", "Marketing", 72000, "2022-06-01", "Meets Expectations"),
        ("Priya Patel", "Engineering", 105000, "2019-11-20", "Exceeds Expectations"),
        ("James Wilson", "Sales", 68000, "2023-01-10", "Meets Expectations"),
        ("Emily Rodriguez", "Finance", 82000, "2020-08-05", "Exceeds Expectations"),
        ("David Kim", "Engineering", 112000, "2018-04-22", "Outstanding"),
        ("Rachel Thompson", "HR", 65000, "2022-09-14", "Meets Expectations"),
        ("Michael Brown", "Operations", 71000, "2021-07-30", "Satisfactory"),
        ("Ana Martinez", "Sales", 77000, "2020-02-18", "Exceeds Expectations"),
        ("Chris Lee", "Marketing", 69000, "2023-05-25", "Meets Expectations"),
        ("Jessica Wang", "Finance", 88000, "2019-06-12", "Outstanding"),
        ("Robert Taylor", "Engineering", 98000, "2020-10-03", "Exceeds Expectations"),
        ("Sophia Nguyen", "Legal", 91000, "2018-12-08", "Outstanding"),
        ("Daniel Garcia", "Support", 58000, "2023-03-20", "Meets Expectations"),
        ("Olivia Anderson", "HR", 63000, "2022-01-15", "Satisfactory"),
        ("William Davis", "Sales", 74000, "2021-04-28", "Exceeds Expectations"),
        ("Emma Thomas", "Engineering", 107000, "2019-09-17", "Outstanding"),
        ("Alexander White", "Operations", 66000, "2022-11-05", "Meets Expectations"),
        ("Isabella Harris", "Marketing", 71000, "2020-05-22", "Meets Expectations"),
        ("Benjamin Clark", "Finance", 85000, "2021-08-14", "Exceeds Expectations"),
        ("Mia Lewis", "Engineering", 99000, "2020-01-07", "Exceeds Expectations"),
        ("Ethan Robinson", "Sales", 72000, "2022-07-19", "Satisfactory"),
        ("Charlotte Walker", "Legal", 87000, "2019-03-25", "Meets Expectations"),
        ("Liam Young", "Support", 56000, "2023-06-30", "Needs Improvement"),
        ("Amelia King", "HR", 67000, "2021-12-11", "Meets Expectations"),
        ("Noah Wright", "Engineering", 115000, "2017-10-02", "Outstanding"),
        ("Harper Lopez", "Marketing", 73000, "2022-04-08", "Exceeds Expectations"),
        ("Mason Hill", "Operations", 69000, "2021-02-27", "Meets Expectations"),
        ("Evelyn Scott", "Finance", 79000, "2020-09-16", "Satisfactory"),
        ("Logan Green", "Sales", 76000, "2019-07-04", "Exceeds Expectations"),
        ("Avery Adams", "Engineering", 102000, "2020-06-20", "Outstanding"),
        ("Ella Baker", "Support", 55000, "2023-08-15", "Meets Expectations"),
        ("Jackson Gonzalez", "Legal", 89000, "2018-11-28", "Exceeds Expectations"),
        ("Scarlett Nelson", "HR", 64000, "2022-03-10", "Meets Expectations"),
        ("Lucas Carter", "Marketing", 70000, "2021-10-22", "Satisfactory"),
        ("Grace Mitchell", "Engineering", 96000, "2020-04-14", "Exceeds Expectations"),
        ("Aiden Perez", "Sales", 71000, "2022-08-06", "Meets Expectations"),
        ("Chloe Roberts", "Finance", 83000, "2019-05-19", "Outstanding"),
        ("Henry Turner", "Operations", 67000, "2021-01-31", "Meets Expectations"),
        ("Lily Phillips", "Engineering", 108000, "2018-08-24", "Outstanding"),
        ("Sebastian Campbell", "Support", 57000, "2023-04-12", "Needs Improvement"),
        ("Zoey Parker", "Legal", 86000, "2020-12-07", "Exceeds Expectations"),
        ("Jack Evans", "HR", 66000, "2022-06-18", "Meets Expectations"),
        ("Penelope Edwards", "Marketing", 74000, "2021-05-09", "Exceeds Expectations"),
        ("Owen Collins", "Sales", 69000, "2020-03-26", "Satisfactory"),
        ("Layla Stewart", "Engineering", 101000, "2019-10-15", "Outstanding"),
        ("Gabriel Sanchez", "Finance", 80000, "2022-02-04", "Meets Expectations"),
        ("Riley Morris", "Operations", 68000, "2021-09-23", "Meets Expectations"),
        ("Nora Rogers", "Engineering", 94000, "2020-07-11", "Exceeds Expectations"),
        ("Carter Reed", "Sales", 75000, "2019-01-29", "Exceeds Expectations"),
        ("Hannah Cook", "Support", 59000, "2023-07-20", "Meets Expectations"),
        ("Dylan Morgan", "Legal", 88000, "2018-06-14", "Outstanding"),
        ("Stella Bell", "HR", 62000, "2022-10-31", "Satisfactory"),
        ("Leo Murphy", "Marketing", 72000, "2021-03-08", "Meets Expectations"),
        ("Aurora Bailey", "Engineering", 103000, "2019-12-22", "Outstanding"),
        ("Julian Rivera", "Finance", 84000, "2020-11-17", "Exceeds Expectations"),
        ("Violet Cooper", "Operations", 70000, "2022-05-13", "Meets Expectations"),
        ("Mateo Richardson", "Sales", 73000, "2021-06-25", "Meets Expectations"),
        ("Hazel Cox", "Engineering", 97000, "2020-02-09", "Exceeds Expectations"),
        ("Elijah Howard", "Support", 56000, "2023-09-01", "Needs Improvement"),
        ("Luna Ward", "Legal", 90000, "2019-04-30", "Exceeds Expectations"),
        ("Grayson Torres", "HR", 65000, "2022-12-16", "Meets Expectations"),
        ("Savannah Peterson", "Marketing", 71000, "2021-08-03", "Satisfactory"),
        ("Lincoln Gray", "Engineering", 110000, "2018-02-19", "Outstanding"),
        ("Paisley Ramirez", "Finance", 81000, "2020-10-28", "Meets Expectations"),
        ("Jayden James", "Sales", 70000, "2022-04-15", "Meets Expectations"),
        ("Addison Watson", "Operations", 66000, "2021-11-09", "Meets Expectations"),
        ("Caleb Brooks", "Engineering", 100000, "2019-08-06", "Outstanding"),
        ("Brooklyn Kelly", "Support", 58000, "2023-02-22", "Meets Expectations"),
        ("Isaac Sanders", "Legal", 92000, "2018-10-18", "Exceeds Expectations"),
        ("Naomi Price", "HR", 64000, "2022-07-07", "Meets Expectations"),
        ("Ezra Bennett", "Marketing", 75000, "2021-04-01", "Exceeds Expectations"),
        ("Aubrey Wood", "Sales", 68000, "2020-06-14", "Satisfactory"),
        ("Adrian Barnes", "Engineering", 104000, "2019-02-25", "Outstanding"),
        ("Claire Ross", "Finance", 86000, "2020-08-20", "Exceeds Expectations"),
        ("Josiah Henderson", "Operations", 69000, "2022-01-28", "Meets Expectations"),
        ("Willow Coleman", "Engineering", 93000, "2021-05-17", "Exceeds Expectations"),
        ("Cameron Jenkins", "Sales", 71000, "2020-09-30", "Meets Expectations"),
        ("Elena Perry", "Support", 57000, "2023-10-05", "Needs Improvement"),
        ("Thomas Powell", "Legal", 87000, "2019-07-12", "Meets Expectations"),
        ("Bella Long", "HR", 63000, "2022-09-08", "Satisfactory"),
        ("Dominic Patterson", "Marketing", 70000, "2021-12-04", "Meets Expectations"),
        ("Madelyn Hughes", "Finance", 82000, "2020-04-18", "Exceeds Expectations"),
        ("Jeremiah Flores", "Engineering", 106000, "2018-05-11", "Outstanding"),
        ("Aria Washington", "Operations", 67000, "2022-08-23", "Meets Expectations"),
        ("Kai Butler", "Sales", 74000, "2021-01-16", "Exceeds Expectations"),
        ("Eliana Simmons", "Engineering", 98000, "2020-03-07", "Exceeds Expectations"),
        ("Christopher Foster", "Support", 60000, "2023-11-19", "Meets Expectations"),
        ("Kinsley Gonzales", "Legal", 85000, "2019-06-28", "Meets Expectations"),
        ("Jaxon Bryant", "HR", 66000, "2022-02-14", "Meets Expectations"),
        ("Maya Alexander", "Marketing", 73000, "2021-07-22", "Satisfactory"),
        ("Andrew Russell", "Finance", 84000, "2020-12-01", "Exceeds Expectations"),
        ("Emilia Griffin", "Engineering", 111000, "2018-09-15", "Outstanding"),
        ("Miles Diaz", "Sales", 69000, "2022-05-30", "Meets Expectations"),
        ("Gianna Hayes", "Operations", 68000, "2021-10-08", "Meets Expectations"),
        ("Samuel Myers", "Engineering", 95000, "2020-01-24", "Exceeds Expectations"),
        ("Isla Ford", "Support", 55000, "2023-06-11", "Needs Improvement"),
        ("Asher Hamilton", "Legal", 91000, "2019-11-03", "Outstanding"),
        ("Quinn Graham", "HR", 65000, "2022-04-26", "Meets Expectations"),
    ]

    for r, (name, dept, salary, start_date, rating) in enumerate(employees, 2):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=dept)
        ws.cell(row=r, column=3, value=salary)
        ws.cell(row=r, column=3).number_format = '$#,##0'
        ws.cell(row=r, column=4, value=start_date)
        ws.cell(row=r, column=5, value=rating)

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 24

    # --- THE BUG: Conditional formatting on entire column A (A:A = A1:A1048576) ---
    # This is a formula-based rule that highlights names containing "son"
    highlight_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    rule = FormulaRule(
        formula=['ISNUMBER(SEARCH("son",A1))'],
        fill=highlight_fill,
    )
    ws.conditional_formatting.add("A1:A1048576", rule)

    # Also add a duplicate broad rule (another overly broad rule to clean up)
    highlight_fill2 = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    rule2 = FormulaRule(
        formula=['ISNUMBER(SEARCH("son",A1))'],
        fill=highlight_fill2,
    )
    ws.conditional_formatting.add("A1:A1048576", rule2)

    # --- Sheet 2: Departments ---
    ws2 = wb.create_sheet("Departments")
    dept_headers = ["Department", "Head Count", "Budget", "Location"]
    for col, h in enumerate(dept_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    dept_data = [
        ("Engineering", 28, 3200000, "Building A - Floor 3"),
        ("Marketing", 12, 980000, "Building B - Floor 1"),
        ("Sales", 15, 1150000, "Building A - Floor 1"),
        ("Finance", 10, 850000, "Building C - Floor 2"),
        ("HR", 8, 620000, "Building B - Floor 2"),
        ("Operations", 9, 710000, "Building A - Floor 2"),
        ("Legal", 7, 680000, "Building C - Floor 1"),
        ("Support", 10, 540000, "Building D - Floor 1"),
    ]
    for r, (dept, count, budget, loc) in enumerate(dept_data, 2):
        ws2.cell(row=r, column=1, value=dept)
        ws2.cell(row=r, column=2, value=count)
        ws2.cell(row=r, column=3, value=budget)
        ws2.cell(row=r, column=3).number_format = '$#,##0'
        ws2.cell(row=r, column=4, value=loc)

    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 24

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
