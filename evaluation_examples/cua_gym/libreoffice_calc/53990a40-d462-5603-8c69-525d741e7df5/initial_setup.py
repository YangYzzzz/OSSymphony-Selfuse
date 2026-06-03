"""
Initial Setup: Shipping and logistics cost comparison spreadsheet
Task ID: calc_grs_068
Domain: libreoffice_calc

Creates a spreadsheet with shipment data and carrier quotes.
Does NOT include: MIN formulas, conditional formatting, weighted scores, or summary formulas.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_068'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ========== Sheet 1: Shipping Quotes ==========
    ws1 = wb.active
    ws1.title = "Shipping Quotes"

    # Headers
    headers = [
        "Shipment ID", "Weight (lbs)", "Dimensions (in)", "Origin Zone",
        "Dest Zone", "FedEx ($)", "UPS ($)", "USPS ($)", "DHL ($)",
        "Regional ($)", "Cheapest ($)",
        "Transit Days", "Reliability (1-5)", "Tracking (Y/N)",
        "Insurance (Y/N)", "Weighted Score"
    ]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Shipment data: 15 rows of realistic shipping quotes
    shipments = [
        ["SHP-2401", 3.2, "10x8x6", "Zone 2", "Zone 5", 12.95, 13.40, 8.75, 15.20, 10.50,
         None, 3, 4.2, "Y", "N", None],
        ["SHP-2402", 8.7, "14x12x10", "Zone 1", "Zone 7", 24.80, 22.15, 18.90, 28.50, 20.75,
         None, 5, 3.8, "Y", "Y", None],
        ["SHP-2403", 1.5, "8x6x4", "Zone 3", "Zone 3", 7.25, 7.80, 5.60, 9.10, 6.90,
         None, 2, 4.5, "Y", "N", None],
        ["SHP-2404", 15.3, "20x16x14", "Zone 4", "Zone 8", 42.60, 39.90, 35.25, 48.75, 37.10,
         None, 7, 3.5, "N", "Y", None],
        ["SHP-2405", 5.0, "12x10x8", "Zone 2", "Zone 6", 18.30, 16.75, 14.20, 21.40, 15.60,
         None, 4, 4.0, "Y", "N", None],
        ["SHP-2406", 22.8, "24x18x16", "Zone 5", "Zone 9", 58.90, 54.25, 47.80, 65.30, 51.40,
         None, 6, 3.2, "Y", "Y", None],
        ["SHP-2407", 0.8, "6x4x3", "Zone 1", "Zone 2", 4.95, 5.20, 3.85, 6.40, 4.50,
         None, 2, 4.8, "Y", "N", None],
        ["SHP-2408", 11.2, "18x14x12", "Zone 3", "Zone 7", 32.45, 29.80, 26.50, 37.90, 28.15,
         None, 5, 3.9, "Y", "Y", None],
        ["SHP-2409", 6.4, "14x10x8", "Zone 6", "Zone 4", 19.75, 18.40, 15.60, 23.10, 17.25,
         None, 3, 4.1, "Y", "N", None],
        ["SHP-2410", 2.1, "9x7x5", "Zone 2", "Zone 3", 8.50, 9.10, 6.25, 10.80, 7.65,
         None, 2, 4.6, "Y", "N", None],
        ["SHP-2411", 18.5, "22x16x14", "Zone 7", "Zone 9", 52.30, 48.60, 42.15, 59.80, 45.90,
         None, 6, 3.4, "N", "Y", None],
        ["SHP-2412", 4.3, "11x9x7", "Zone 1", "Zone 5", 15.60, 14.25, 11.80, 18.40, 13.20,
         None, 4, 4.3, "Y", "N", None],
        ["SHP-2413", 9.8, "16x12x10", "Zone 4", "Zone 6", 28.90, 26.50, 22.75, 33.60, 24.80,
         None, 4, 3.7, "Y", "Y", None],
        ["SHP-2414", 0.5, "5x4x3", "Zone 2", "Zone 1", 3.80, 4.15, 2.95, 5.20, 3.50,
         None, 1, 4.9, "Y", "N", None],
        ["SHP-2415", 13.6, "19x15x12", "Zone 5", "Zone 8", 38.75, 35.90, 31.40, 44.20, 33.60,
         None, 5, 3.6, "Y", "Y", None],
    ]

    money_fmt = '$#,##0.00'
    center_align = Alignment(horizontal="center", vertical="center")

    for r, row_data in enumerate(shipments, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.border = thin_border
            cell.alignment = center_align
            # Money format for carrier quote columns (F-K) and weighted score (P)
            if c in (6, 7, 8, 9, 10, 11, 16):
                cell.number_format = money_fmt
            # Reliability score format
            if c == 13:
                cell.number_format = '0.0'

    # Set column widths
    col_widths = {
        'A': 14, 'B': 14, 'C': 15, 'D': 13, 'E': 13,
        'F': 12, 'G': 12, 'H': 12, 'I': 12, 'J': 14, 'K': 14,
        'L': 13, 'M': 16, 'N': 14, 'O': 15, 'P': 16,
    }
    for col_letter, width in col_widths.items():
        ws1.column_dimensions[col_letter].width = width

    # Freeze header row
    ws1.freeze_panes = "A2"

    # ========== Sheet 2: Summary ==========
    ws2 = wb.create_sheet("Summary")

    # Title
    ws2.merge_cells("A1:E1")
    title_cell = ws2["A1"]
    title_cell.value = "Carrier Performance Summary"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Summary headers
    summary_headers = ["Carrier", "Win Rate (%)", "Avg Price ($)", "Total Quotes ($)", "Potential Savings ($)"]
    for col, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Carrier names only (no formulas/values - those are the task)
    carriers = ["FedEx", "UPS", "USPS", "DHL", "Regional Carrier"]
    for r, carrier in enumerate(carriers, 4):
        cell = ws2.cell(row=r, column=1, value=carrier)
        cell.border = thin_border
        cell.font = Font(name="Calibri", size=11)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        for c in range(2, 6):
            cell = ws2.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if c == 2:
                cell.number_format = '0.0%'
            else:
                cell.number_format = '$#,##0.00'

    # Total savings row label
    ws2.cell(row=10, column=1, value="Total Potential Savings:").font = Font(
        name="Calibri", size=12, bold=True, color="2F5496"
    )
    ws2.cell(row=10, column=2).number_format = '$#,##0.00'

    # Column widths for summary
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 16
    ws2.column_dimensions['D'].width = 18
    ws2.column_dimensions['E'].width = 20

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
