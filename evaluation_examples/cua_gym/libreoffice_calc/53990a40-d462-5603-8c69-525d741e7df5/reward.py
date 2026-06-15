"""
Reward Script: Shipping and logistics cost comparison in LibreOffice Calc
Task ID: calc_grs_068
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): MIN formulas in K2:K16 (cheapest option)
  Component 2 (0.20): Weighted Score formulas in P2:P16
  Component 3 (0.25): Conditional formatting on carrier quote columns (green for min)
  Component 4 (0.20): Summary sheet formulas (win rate, avg price, total quotes, savings)
  Component 5 (0.10): Summary total potential savings formula in B10
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_068'


def persist_app_state(domain):
    """Try to save any unsaved LibreOffice state via Ctrl+S."""
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print(f"PERSIST: ctrl+s sent for {domain}")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: required sheets exist
    if 'Shipping Quotes' not in wb.sheetnames:
        print("FAIL: 'Shipping Quotes' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Shipping Quotes']

    # =========================================================================
    # Component 1: MIN formulas in K2:K16 (0.25 points)
    # These identify the cheapest carrier for each shipment.
    # Initial has K2:K16 empty; golden has =MIN(F,G,H,I,J) formulas.
    # =========================================================================
    try:
        min_formula_count = 0
        for r in range(2, 17):
            val = ws.cell(row=r, column=11).value  # column K
            if val is not None and isinstance(val, str) and 'MIN' in val.upper():
                # Check it references the carrier columns F-J
                upper_val = val.upper().replace(" ", "")
                if 'F' in upper_val and 'J' in upper_val:
                    min_formula_count += 1
        if min_formula_count == 15:
            print(f"PASS: Component 1 -- All 15 MIN formulas found in K2:K16 (0.25 pts)")
            total_score += 0.25
        elif min_formula_count >= 10:
            partial = 0.25 * (min_formula_count / 15)
            print(f"PARTIAL: Component 1 -- {min_formula_count}/15 MIN formulas found ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 -- Only {min_formula_count}/15 MIN formulas found in K2:K16")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # =========================================================================
    # Component 2: Weighted Score formulas in P2:P16 (0.20 points)
    # These combine cost and service factors. Initial has P2:P16 empty.
    # Golden uses formulas referencing K column (cost) and M column (reliability).
    # =========================================================================
    try:
        weighted_formula_count = 0
        for r in range(2, 17):
            val = ws.cell(row=r, column=16).value  # column P
            if val is not None and isinstance(val, str):
                upper_val = val.upper().replace(" ", "")
                # Must be a formula that references cost (K or F-J) and reliability (M)
                if upper_val.startswith('=') and 'M' in upper_val:
                    weighted_formula_count += 1
        if weighted_formula_count == 15:
            print(f"PASS: Component 2 -- All 15 Weighted Score formulas found in P2:P16 (0.20 pts)")
            total_score += 0.20
        elif weighted_formula_count >= 10:
            partial = 0.20 * (weighted_formula_count / 15)
            print(f"PARTIAL: Component 2 -- {weighted_formula_count}/15 weighted formulas ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- Only {weighted_formula_count}/15 Weighted Score formulas in P2:P16")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # =========================================================================
    # Component 3: Conditional formatting with green highlight for min cost (0.25 points)
    # Initial has 0 CF rules; golden has CF rules on carrier columns F-J
    # that highlight when a cell equals the MIN of the row in green.
    # =========================================================================
    try:
        cf_rules = ws.conditional_formatting
        # Collect all CF ranges that apply to carrier columns (F=6, G=7, H=8, I=9, J=10)
        carrier_cf_count = 0
        has_green_fill = False
        has_min_formula = False

        for cf in cf_rules:
            cf_range_str = str(cf)
            for rule in cf.rules:
                # Check if the formula references MIN
                formula_list = getattr(rule, 'formula', None) or []
                for f in formula_list:
                    if 'MIN' in f.upper():
                        has_min_formula = True

                # Check for green fill
                if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                    try:
                        fg_rgb = rule.dxf.fill.fgColor.rgb
                        if fg_rgb is not None:
                            fg_upper = str(fg_rgb).upper()
                            # Green variants: FF92D050, FF00FF00, FF00B050, etc.
                            # Check if it's a greenish color (G component > R and B)
                            if len(fg_upper) == 8:
                                r_val = int(fg_upper[2:4], 16)
                                g_val = int(fg_upper[4:6], 16)
                                b_val = int(fg_upper[6:8], 16)
                                if g_val > r_val and g_val > b_val:
                                    has_green_fill = True
                    except Exception:
                        pass

            # Check if the range covers any carrier columns
            for col_letter in ['F', 'G', 'H', 'I', 'J']:
                if col_letter in cf_range_str:
                    carrier_cf_count += 1
                    break

        if carrier_cf_count >= 3 and has_green_fill and has_min_formula:
            print(f"PASS: Component 3 -- Conditional formatting with green MIN highlight on carrier columns (0.25 pts)")
            total_score += 0.25
        elif carrier_cf_count >= 1 and (has_green_fill or has_min_formula):
            print(f"PARTIAL: Component 3 -- Some CF found: {carrier_cf_count} carrier ranges, green={has_green_fill}, min_formula={has_min_formula} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 -- No valid conditional formatting found. Carrier CF ranges: {carrier_cf_count}, green: {has_green_fill}, min: {has_min_formula}")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # =========================================================================
    # Component 4: Summary sheet formulas for carrier metrics (0.20 points)
    # Initial Summary has only headers/labels, no formulas.
    # Golden has formulas in B4:E8 for each carrier's win rate, avg price, total, savings.
    # =========================================================================
    try:
        if 'Summary' not in wb.sheetnames:
            print("FAIL: Component 4 -- 'Summary' sheet not found")
        else:
            ws_sum = wb['Summary']
            formula_count = 0
            # Check cells B4:E8 (5 carriers x 4 metrics = 20 cells)
            for r in range(4, 9):
                for c in range(2, 6):  # columns B=2, C=3, D=4, E=5
                    val = ws_sum.cell(row=r, column=c).value
                    if val is not None and isinstance(val, str) and val.startswith('='):
                        formula_count += 1

            if formula_count >= 15:
                print(f"PASS: Component 4 -- {formula_count}/20 Summary formulas found in B4:E8 (0.20 pts)")
                total_score += 0.20
            elif formula_count >= 8:
                partial = 0.20 * (formula_count / 20)
                print(f"PARTIAL: Component 4 -- {formula_count}/20 Summary formulas ({partial:.2f} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 4 -- Only {formula_count}/20 Summary formulas found in B4:E8")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # =========================================================================
    # Component 5: Total Potential Savings formula (0.10 points)
    # Initial has no formula in the total savings area; golden has one.
    # =========================================================================
    try:
        if 'Summary' not in wb.sheetnames:
            print("FAIL: Component 5 -- 'Summary' sheet not found")
        else:
            ws_sum = wb['Summary']
            # Check row 10 area for a savings total formula
            savings_found = False
            for r in range(9, 12):
                for c in range(1, 6):
                    val = ws_sum.cell(row=r, column=c).value
                    if val is not None and isinstance(val, str) and val.startswith('='):
                        # Should reference the Shipping Quotes sheet or savings columns
                        if 'SUM' in val.upper() or 'AVERAGE' in val.upper() or 'Shipping' in val:
                            savings_found = True
                            print(f"  Found savings formula at {ws_sum.cell(row=r, column=c).coordinate}: {val}")
                            break
                if savings_found:
                    break

            if savings_found:
                print(f"PASS: Component 5 -- Total Potential Savings formula found (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 5 -- No Total Potential Savings formula found in Summary")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
