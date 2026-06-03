"""
Initial Setup: Create enrollment data for pivot table task
Task ID: calc_pivot_067
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_067'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Enrollments'

    # --- Headers ---
    headers = ['ID', 'Department', 'CourseLevel', 'StudentID', 'Semester']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Department / CourseLevel distribution ---
    # We need exactly 500 rows with a specific distribution.
    # The ground truth says Engineering/100 = 42 (8.4%).
    # We'll build a deterministic count matrix:
    #
    #              100   200   300   400   Total
    # Engineering   42    30    18    10    100
    # Business      35    28    22    15    100
    # Sciences      30    32    25    13    100
    # Humanities    28    26    30    16    100
    # Health Sci    25    24    27    24    100
    # Total        160   140   122    78    500

    departments = ['Engineering', 'Business', 'Sciences', 'Humanities', 'Health Sciences']
    course_levels = [100, 200, 300, 400]

    counts = {
        'Engineering':      [42, 30, 18, 10],
        'Business':         [35, 28, 22, 15],
        'Sciences':         [30, 32, 25, 13],
        'Humanities':       [28, 26, 30, 16],
        'Health Sciences':  [25, 24, 27, 24],
    }

    semesters = ['Fall 2024', 'Spring 2025', 'Fall 2025', 'Spring 2026']

    # First names and last names for realistic StudentIDs
    first_names = [
        'Sarah', 'Marcus', 'Emily', 'James', 'Priya', 'Carlos', 'Wei',
        'Fatima', 'Alex', 'Nadia', 'Thomas', 'Yuki', 'Rachel', 'David',
        'Aisha', 'Michael', 'Sophia', 'Daniel', 'Olivia', 'Ryan',
        'Isabella', 'Nathan', 'Maya', 'Liam', 'Zara', 'Ethan', 'Hannah',
        'Omar', 'Chloe', 'Jayden', 'Amara', 'Lucas', 'Mia', 'Kevin',
        'Sofia', 'Brandon', 'Elena', 'Jordan', 'Lily', 'Victor'
    ]
    last_names = [
        'Chen', 'Johnson', 'Patel', 'Kim', 'Garcia', 'Williams', 'Zhang',
        'Ahmed', 'Brown', 'Singh', 'Martinez', 'Lee', 'Anderson', 'Taylor',
        'Wilson', 'Moore', 'Jackson', 'White', 'Harris', 'Clark',
        'Lewis', 'Robinson', 'Walker', 'Young', 'Allen', 'King', 'Wright',
        'Scott', 'Adams', 'Baker', 'Gonzalez', 'Nelson', 'Carter', 'Mitchell',
        'Perez', 'Roberts', 'Turner', 'Phillips', 'Campbell', 'Parker'
    ]

    # Build all 500 enrollment records
    records = []
    for dept in departments:
        for level_idx, level in enumerate(course_levels):
            count = counts[dept][level_idx]
            for _ in range(count):
                fn = random.choice(first_names)
                ln = random.choice(last_names)
                student_id = f'S{random.randint(10000, 99999)}'
                semester = random.choice(semesters)
                records.append([dept, level, student_id, semester])

    # Shuffle for realism
    random.shuffle(records)

    # Write data rows
    for row_idx, (dept, level, sid, sem) in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)  # ID: 1-500
        ws.cell(row=row_idx, column=2, value=dept)
        ws.cell(row=row_idx, column=3, value=level)
        ws.cell(row=row_idx, column=4, value=sid)
        ws.cell(row=row_idx, column=5, value=sem)

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 16

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Total records: {len(records)}')

    # Verify counts
    from collections import Counter
    dept_level_counts = Counter()
    for dept, level, sid, sem in records:
        dept_level_counts[(dept, level)] += 1
    print(f'Engineering/100 count: {dept_level_counts[("Engineering", 100)]}')
    print(f'Total unique (dept,level) combos: {len(dept_level_counts)}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
