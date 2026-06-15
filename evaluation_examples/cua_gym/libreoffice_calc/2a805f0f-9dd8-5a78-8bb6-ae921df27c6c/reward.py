"""
Reward Script: Create pivot table from enrollment data as % of grand total
Task ID: calc_pivot_067
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): PivotTable sheet exists with data
  Component 2 (0.25): Correct structure (5 dept rows + Grand Total, 4 course level cols + Grand Total)
  Component 3 (0.30): Values are percentages summing to 1.0 (100%)
  Component 4 (0.20): Specific cell values match expected percentages
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_067'

# Expected departments and course levels from the task context
EXPECTED_DEPARTMENTS = {'Engineering', 'Business', 'Sciences', 'Humanities', 'Health Sciences'}
EXPECTED_COURSE_LEVELS = {100, 200, 300, 400}

# Expected values from task context (as fractions of 500 total)
# Engineering/100 = 42/500 = 0.084, etc.
EXPECTED_VALUES = {
    ('Engineering', 100): 0.084,
    ('Engineering', 200): 0.06,
    ('Engineering', 300): 0.036,
    ('Engineering', 400): 0.02,
    ('Business', 100): 0.07,
    ('Business', 200): 0.056,
    ('Business', 300): 0.044,
    ('Business', 400): 0.03,
    ('Sciences', 100): 0.06,
    ('Sciences', 200): 0.064,
    ('Sciences', 300): 0.05,
    ('Sciences', 400): 0.026,
    ('Humanities', 100): 0.056,
    ('Humanities', 200): 0.052,
    ('Humanities', 300): 0.06,
    ('Humanities', 400): 0.032,
    ('Health Sciences', 100): 0.05,
    ('Health Sciences', 200): 0.048,
    ('Health Sciences', 300): 0.054,
    ('Health Sciences', 400): 0.048,
}


def find_pivot_sheet(wb):
    """Find a sheet that looks like a pivot table (not the raw Enrollments data).
    Returns the worksheet or None."""
    for sn in wb.sheetnames:
        if sn == 'Enrollments':
            continue
        ws = wb[sn]
        # Check if it has some data that looks like a pivot table
        if ws.max_row >= 2 and ws.max_column >= 2:
            return ws
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Pivot table sheet exists with data (0.25 points)
    # This FAILS on initial (only has 'Enrollments'), PASSES on golden (has 'PivotTable')
    try:
        pivot_ws = find_pivot_sheet(wb)
        if pivot_ws is not None and pivot_ws.max_row >= 6 and pivot_ws.max_column >= 5:
            print(f"PASS: Component 1 — Pivot table sheet '{pivot_ws.title}' exists with data ({pivot_ws.max_row} rows, {pivot_ws.max_column} cols) (0.25 pts)")
            total_score += 0.25
        else:
            if pivot_ws is None:
                print("FAIL: Component 1 — No pivot table sheet found (only 'Enrollments')")
            else:
                print(f"FAIL: Component 1 — Pivot sheet '{pivot_ws.title}' too small: {pivot_ws.max_row} rows, {pivot_ws.max_column} cols")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print(f"REWARD: {total_score}")
        return total_score

    # Component 2: Correct structure — 5 departments + Grand Total row, 4 course levels + Grand Total col (0.25 points)
    try:
        # Read header row to find course level columns
        header_row = {}
        for col in range(1, pivot_ws.max_column + 1):
            val = pivot_ws.cell(row=1, column=col).value
            if val is not None:
                header_row[col] = val

        # Read department names from column A (rows 2+)
        found_depts = set()
        grand_total_row_count = 0
        for row in range(2, pivot_ws.max_row + 1):
            val = pivot_ws.cell(row=row, column=1).value
            if val is not None:
                val_str = str(val).strip()
                if val_str.lower() == 'grand total':
                    grand_total_row_count += 1
                else:
                    found_depts.add(val_str)
        has_grand_total_row = grand_total_row_count > 0

        # Check course level columns (look for 100, 200, 300, 400 in header)
        found_levels = set()
        grand_total_col_count = 0
        for col, val in header_row.items():
            if col == 1:
                continue  # skip the row label column
            try:
                int_val = int(val)
                if int_val in EXPECTED_COURSE_LEVELS:
                    found_levels.add(int_val)
            except (ValueError, TypeError):
                if str(val).strip().lower() == 'grand total':
                    grand_total_col_count += 1
        has_grand_total_col = grand_total_col_count > 0

        dept_match = len(found_depts.intersection(EXPECTED_DEPARTMENTS)) >= 5
        level_match = len(found_levels) >= 4

        structure_score = 0.0
        if dept_match:
            structure_score += 0.08
        if level_match:
            structure_score += 0.08
        if has_grand_total_row:
            structure_score += 0.045
        if has_grand_total_col:
            structure_score += 0.045

        if structure_score > 0:
            print(f"PASS: Component 2 — Structure check: depts={found_depts}, levels={found_levels}, "
                  f"grand_total_row={has_grand_total_row}, grand_total_col={has_grand_total_col} ({structure_score:.3f} pts)")
            total_score += structure_score
        else:
            print(f"FAIL: Component 2 — Structure mismatch: found depts={found_depts}, levels={found_levels}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Values are percentages that sum to 1.0 (0.30 points)
    # Each cell should be a fraction (0 < val < 1) and Grand Total should be 1.0
    try:
        # Build a map of (dept, level) -> value
        # First map col index to course level
        col_to_level = {}
        grand_total_col_idx = None
        for col in range(2, pivot_ws.max_column + 1):
            hval = pivot_ws.cell(row=1, column=col).value
            if hval is not None:
                try:
                    col_to_level[col] = int(hval)
                except (ValueError, TypeError):
                    if str(hval).strip().lower() == 'grand total':
                        grand_total_col_idx = col

        # Check that data cells are percentages (between 0 and 1)
        all_data_values = []
        grand_total_value = None
        for row in range(2, pivot_ws.max_row + 1):
            row_label = pivot_ws.cell(row=row, column=1).value
            if row_label is None:
                continue
            row_label_str = str(row_label).strip()

            for col in range(2, pivot_ws.max_column + 1):
                cell_val = pivot_ws.cell(row=row, column=col).value
                if cell_val is None:
                    continue

                try:
                    fval = float(cell_val)
                except (ValueError, TypeError):
                    continue

                if row_label_str.lower() == 'grand total' and col == grand_total_col_idx:
                    grand_total_value = fval
                elif row_label_str.lower() != 'grand total' and col in col_to_level:
                    all_data_values.append(fval)

        pct_score = 0.0
        # Sub-check: all data values are between 0 and 1 (exclusive)
        if all_data_values and all(0 < v < 1 for v in all_data_values):
            pct_score += 0.10
            print(f"  Sub-check 3a: All {len(all_data_values)} data values are in (0, 1) range")
        else:
            print(f"  Sub-check 3a FAIL: Data values not in percentage range. Sample: {all_data_values[:5]}")

        # Sub-check: data values sum close to 1.0
        data_sum = sum(all_data_values) if all_data_values else 0
        if abs(data_sum - 1.0) < 0.02:
            pct_score += 0.10
            print(f"  Sub-check 3b: Data values sum to {data_sum:.4f} (close to 1.0)")
        else:
            print(f"  Sub-check 3b FAIL: Data values sum to {data_sum:.4f}, expected ~1.0")

        # Sub-check: Grand Total cell == 1.0
        if grand_total_value is not None and abs(grand_total_value - 1.0) < 0.01:
            pct_score += 0.10
            print(f"  Sub-check 3c: Grand Total = {grand_total_value} (correct)")
        else:
            print(f"  Sub-check 3c FAIL: Grand Total = {grand_total_value}, expected 1.0")

        if pct_score > 0:
            print(f"PASS: Component 3 — Percentage verification ({pct_score:.2f} pts)")
            total_score += pct_score
        else:
            print("FAIL: Component 3 — Values do not represent valid percentages of grand total")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Specific cell values match expected percentages (0.20 points)
    # Spot-check key values from the task context
    try:
        # Build lookup: (dept_name, course_level) -> value from the pivot sheet
        actual_values = {}
        for row in range(2, pivot_ws.max_row + 1):
            row_label = pivot_ws.cell(row=row, column=1).value
            if row_label is None:
                continue
            row_label_str = str(row_label).strip()
            if row_label_str.lower() == 'grand total':
                continue

            for col, level in col_to_level.items():
                cell_val = pivot_ws.cell(row=row, column=col).value
                if cell_val is not None:
                    try:
                        actual_values[(row_label_str, level)] = float(cell_val)
                    except (ValueError, TypeError):
                        pass

        # Check a selection of expected values
        spot_checks = [
            ('Engineering', 100, 0.084),  # context says 8.4% = 42/500
            ('Engineering', 200, 0.06),
            ('Business', 100, 0.07),
            ('Health Sciences', 400, 0.048),
            ('Sciences', 300, 0.05),
        ]

        matches = 0
        for dept, level, expected_val in spot_checks:
            actual = actual_values.get((dept, level))
            if actual is not None and abs(actual - expected_val) < 0.005:
                matches += 1
                print(f"  Spot-check ({dept}, {level}): expected {expected_val}, got {actual} - MATCH")
            else:
                print(f"  Spot-check ({dept}, {level}): expected {expected_val}, got {actual} - MISMATCH")

        value_score = 0.20 * (matches / len(spot_checks))
        if value_score > 0:
            print(f"PASS: Component 4 — {matches}/{len(spot_checks)} spot checks passed ({value_score:.2f} pts)")
            total_score += value_score
        else:
            print(f"FAIL: Component 4 — 0/{len(spot_checks)} spot checks passed")

    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
