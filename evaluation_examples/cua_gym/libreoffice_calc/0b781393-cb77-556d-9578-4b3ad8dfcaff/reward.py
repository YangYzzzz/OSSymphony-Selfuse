"""
Reward Script: Group SalesAmount into ranges of 1000 and show count of transactions per range
Task ID: calc_pivot_060
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20): PivotSummary sheet exists
  Component 2 (0.30): Correct range labels (0-999 through 4000-4999)
  Component 3 (0.40): Correct count values per range (85, 92, 68, 35, 20)
  Component 4 (0.10): Grand Total row present with value 300
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_060'

# Expected pivot data from task context ground truth
EXPECTED_RANGES = {
    '0-999': 85,
    '1000-1999': 92,
    '2000-2999': 68,
    '3000-3999': 35,
    '4000-4999': 20,
}
EXPECTED_GRAND_TOTAL = 300


def persist_app_state(domain):
    """Best-effort save of any open LibreOffice document."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def find_pivot_sheet(wb):
    """
    Find a sheet that contains the pivot/summary data.
    Look for any sheet other than 'TxnData' that has grouped SalesAmount ranges.
    Also search TxnData itself in case the pivot was placed there.
    """
    # First, check for any non-TxnData sheet
    for sn in wb.sheetnames:
        if sn.lower() != 'txndata':
            return wb[sn], sn
    return None, None


def find_data_region(ws):
    """
    Scan the sheet for the pivot table data region.
    Returns a dict mapping range_label -> count_value, plus grand_total if found.
    We look for cells containing patterns like '0-999', '1000-1999', etc.
    """
    range_data = {}
    grand_total = None

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            val = cell.value
            if val is None:
                continue
            val_str = str(val).strip()

            # Check if this looks like a range label (e.g. "0-999", "1000-1999")
            for expected_range in EXPECTED_RANGES:
                if val_str == expected_range:
                    # Look for the count value in the next column(s) on the same row
                    for offset in range(1, 5):
                        neighbor = ws.cell(row=cell.row, column=cell.column + offset).value
                        if neighbor is not None:
                            try:
                                range_data[expected_range] = int(float(neighbor))
                            except (ValueError, TypeError):
                                pass
                            break
                    break

            # Check for grand total row
            if 'grand' in val_str.lower() and 'total' in val_str.lower():
                for offset in range(1, 5):
                    neighbor = ws.cell(row=cell.row, column=cell.column + offset).value
                    if neighbor is not None:
                        try:
                            grand_total = int(float(neighbor))
                        except (ValueError, TypeError):
                            pass
                        break

    return range_data, grand_total


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: PivotSummary (or equivalent) sheet exists (0.20 points)
    # This is a task-introduced change: initial has only TxnData
    try:
        pivot_ws, pivot_name = find_pivot_sheet(wb)
        if pivot_ws is not None:
            print(f"PASS: Component 1 — Pivot sheet '{pivot_name}' exists (0.20 pts)")
            total_score += 0.20
        else:
            print("FAIL: Component 1 — No pivot/summary sheet found (only TxnData exists)")
            # If no pivot sheet at all, no point checking further
            final_score = min(total_score, 1.0)
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {final_score}")
            return final_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Correct range labels present (0.30 points)
    # Each of the 5 expected ranges contributes 0.06 points
    try:
        range_data, grand_total_val = find_data_region(pivot_ws)
        labels_found = 0
        for rng in EXPECTED_RANGES:
            if rng in range_data:
                labels_found += 1
                print(f"  Found range label: {rng}")
            else:
                print(f"  Missing range label: {rng}")

        label_score = (labels_found / len(EXPECTED_RANGES)) * 0.30
        if labels_found == len(EXPECTED_RANGES):
            print(f"PASS: Component 2 — All 5 range labels found ({label_score:.2f} pts)")
            total_score += label_score
        elif labels_found > 0:
            print(f"PARTIAL: Component 2 — {labels_found}/5 range labels found ({label_score:.2f} pts)")
            total_score += label_score
        else:
            print(f"FAIL: Component 2 — No expected range labels found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Correct count values per range (0.40 points)
    # Each of the 5 expected values contributes 0.08 points
    try:
        values_correct = 0
        for rng, expected_count in EXPECTED_RANGES.items():
            actual = range_data.get(rng)
            if actual is not None and actual == expected_count:
                values_correct += 1
                print(f"  Count correct for {rng}: {actual} == {expected_count}")
            else:
                print(f"  Count wrong for {rng}: found {actual}, expected {expected_count}")

        value_score = (values_correct / len(EXPECTED_RANGES)) * 0.40
        if values_correct == len(EXPECTED_RANGES):
            print(f"PASS: Component 3 — All 5 count values correct ({value_score:.2f} pts)")
            total_score += value_score
        elif values_correct > 0:
            print(f"PARTIAL: Component 3 — {values_correct}/5 count values correct ({value_score:.2f} pts)")
            total_score += value_score
        else:
            print(f"FAIL: Component 3 — No count values match expected")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Grand Total row with value 300 (0.10 points)
    try:
        if grand_total_val is not None and grand_total_val == EXPECTED_GRAND_TOTAL:
            print(f"PASS: Component 4 — Grand Total = {grand_total_val} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 — Grand Total: found {grand_total_val}, expected {EXPECTED_GRAND_TOTAL}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
