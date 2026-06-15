"""
Initial Setup: Create transaction data spreadsheet for pivot table grouping task
Task ID: calc_pivot_060
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_060'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'TxnData'

    # --- Headers ---
    headers = ['TxnID', 'Customer', 'SalesAmount', 'Date']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_side = Side(style='thin', color='000000')
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # --- Generate 300 transactions with exact distribution ---
    # Ground truth distribution:
    #   0-999:     85 transactions
    #   1000-1999: 92 transactions
    #   2000-2999: 68 transactions
    #   3000-3999: 35 transactions
    #   4000-4999: 20 transactions
    #   Total: 300

    distribution = [
        (50, 999, 85),
        (1000, 1999, 92),
        (2000, 2999, 68),
        (3000, 3999, 35),
        (4000, 4999, 20),
    ]

    amounts = []
    for low, high, count in distribution:
        for _ in range(count):
            amounts.append(round(random.uniform(low, high), 2))

    assert len(amounts) == 300, f"Expected 300 amounts, got {len(amounts)}"

    # Shuffle so they're not in order by range
    random.shuffle(amounts)

    # Customer names pool
    first_names = [
        'Sarah', 'Marcus', 'Elena', 'James', 'Priya', 'David', 'Mei',
        'Carlos', 'Amara', 'Liam', 'Fatima', 'Noah', 'Yuki', 'Ethan',
        'Zara', 'Oliver', 'Anya', 'Lucas', 'Nadia', 'Ryan', 'Sofia',
        'Tyler', 'Aisha', 'Benjamin', 'Clara', 'Daniel', 'Eva', 'Frank',
        'Grace', 'Henry', 'Isabella', 'Jack', 'Kelly', 'Leo', 'Maria',
        'Nathan', 'Olivia', 'Patrick', 'Quinn', 'Rachel',
    ]
    last_names = [
        'Chen', 'Johnson', 'Petrov', 'Williams', 'Sharma', 'Kim',
        'Anderson', 'Lopez', 'Mueller', 'Nakamura', 'Brown', 'Garcia',
        'Singh', 'Tanaka', 'Wilson', 'Martinez', 'Lee', 'Robinson',
        'Clark', 'Lewis', 'Walker', 'Hall', 'Young', 'Allen', 'King',
        'Wright', 'Scott', 'Torres', 'Adams', 'Baker',
    ]

    # Dates spanning 2024-2025
    start_year = 2024
    end_year = 2025

    data_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )

    for i in range(300):
        row = i + 2
        txn_id = i + 1
        customer = f'{random.choice(first_names)} {random.choice(last_names)}'
        amount = amounts[i]
        year = random.choice([start_year, end_year])
        month = random.randint(1, 12)
        day = random.randint(1, 28)
        date_str = f'{year}-{month:02d}-{day:02d}'

        ws.cell(row=row, column=1, value=txn_id).border = data_border
        ws.cell(row=row, column=2, value=customer).border = data_border
        c = ws.cell(row=row, column=3, value=amount)
        c.number_format = '#,##0.00'
        c.border = data_border
        ws.cell(row=row, column=4, value=date_str).border = data_border

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
