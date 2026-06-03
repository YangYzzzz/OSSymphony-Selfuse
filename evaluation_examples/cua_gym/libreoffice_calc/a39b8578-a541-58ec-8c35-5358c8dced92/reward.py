"""
Reward Script: Calculate lead conversion rate from the number of leads and won deals.
Task ID: calc_sales_023
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): E2:E6 contain formulas (not empty/None) — 0.1 per cell
  Component 2 (0.3): Formulas are correct division =Dn/Bn — 0.06 per cell
  Component 3 (0.2): E2:E6 formatted as percentage — 0.04 per cell
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_023'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet exists
    if 'Conversion' not in wb.sheetnames:
        print("FAIL: Sheet 'Conversion' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Conversion']

    # Target cells and expected formulas
    # E2=D2/B2, E3=D3/B3, E4=D4/B4, E5=D5/B5, E6=D6/B6
    cells = [2, 3, 4, 5, 6]

    # Component 1: E2:E6 contain formulas (not empty) — 0.1 per cell, total 0.5
    print("\n--- Component 1: Formula presence in E2:E6 (0.5 pts) ---")
    for row in cells:
        coord = f"E{row}"
        try:
            val = ws[coord].value
            if val is not None and isinstance(val, str) and val.startswith('='):
                print(f"  PASS: {coord} has formula: {val} (+0.1)")
                total_score += 0.1
            elif val is not None and isinstance(val, (int, float)):
                # Agent may have entered a computed value instead of formula
                print(f"  PASS: {coord} has numeric value: {val} (+0.1)")
                total_score += 0.1
            else:
                print(f"  FAIL: {coord} is empty or not a formula/number: {val!r}")
        except Exception as e:
            print(f"  ERROR: {coord}: {e}")

    # Component 2: Formulas are correct division =Dn/Bn — 0.06 per cell, total 0.3
    # Accept both formula form and correct numeric value
    # Ground truth values: E2=0.1, E3~0.10526, E4~0.12857, E5~0.13636, E6~0.10769
    expected_values = {
        2: 12 / 120,    # 0.1
        3: 10 / 95,     # 0.10526...
        4: 18 / 140,    # 0.12857...
        5: 15 / 110,    # 0.13636...
        6: 14 / 130,    # 0.10769...
    }

    print("\n--- Component 2: Correct conversion formula =Dn/Bn (0.3 pts) ---")
    for row in cells:
        coord = f"E{row}"
        try:
            val = ws[coord].value
            if isinstance(val, str):
                # Check formula pattern: =D{row}/B{row}
                normalized = val.upper().replace(" ", "")
                expected_formula = f"=D{row}/B{row}"
                if normalized == expected_formula:
                    print(f"  PASS: {coord} formula correct: {val} (+0.06)")
                    total_score += 0.06
                else:
                    print(f"  FAIL: {coord} formula mismatch: expected {expected_formula}, got {val}")
            elif isinstance(val, (int, float)):
                # Check numeric value with tolerance
                expected = expected_values[row]
                if abs(float(val) - expected) < 0.005:
                    print(f"  PASS: {coord} numeric value correct: {val} ~ {expected:.5f} (+0.06)")
                    total_score += 0.06
                else:
                    print(f"  FAIL: {coord} numeric value wrong: {val}, expected ~{expected:.5f}")
            else:
                print(f"  FAIL: {coord} has no formula or value: {val!r}")
        except Exception as e:
            print(f"  ERROR: {coord}: {e}")

    # Component 3: Percentage number format on E2:E6 — 0.04 per cell, total 0.2
    print("\n--- Component 3: Percentage format on E2:E6 (0.2 pts) ---")
    for row in cells:
        coord = f"E{row}"
        try:
            nf = ws[coord].number_format
            # Accept any format containing '%'
            if nf and '%' in nf:
                print(f"  PASS: {coord} formatted as percentage: {nf!r} (+0.04)")
                total_score += 0.04
            else:
                print(f"  FAIL: {coord} not formatted as percentage: {nf!r}")
        except Exception as e:
            print(f"  ERROR: {coord}: {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
