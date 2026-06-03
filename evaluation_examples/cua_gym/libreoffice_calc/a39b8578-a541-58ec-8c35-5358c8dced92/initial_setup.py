"""
Initial Setup: Calculate lead conversion rate from the number of leads and won deals.
Task ID: calc_sales_023
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_023'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Conversion ---
    ws = wb.active
    ws.title = 'Conversion'

    # Headers
    headers = ['Month', 'Leads', 'Qualified', 'Won', 'Conversion Rate']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    data = [
        ['Jan', 120, 48, 12],
        ['Feb', 95, 38, 10],
        ['Mar', 140, 56, 18],
        ['Apr', 110, 44, 15],
        ['May', 130, 52, 14],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 1:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="right")

    # E2:E6 intentionally left empty - the agent must fill these in
    for r in range(2, 7):
        cell = ws.cell(row=r, column=5)
        cell.border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 18

    # --- Sheet: Pipeline Summary (extra context sheet) ---
    ws2 = wb.create_sheet('Pipeline Summary')
    ws2['A1'] = 'Sales Pipeline Overview'
    ws2['A1'].font = Font(name='Calibri', size=14, bold=True)
    ws2['A3'] = 'Region'
    ws2['B3'] = 'Total Leads (Jan-May)'
    ws2['C3'] = 'Total Won (Jan-May)'
    ws2['A3'].font = Font(bold=True)
    ws2['B3'].font = Font(bold=True)
    ws2['C3'].font = Font(bold=True)

    regions = [
        ['North America', 345, 42],
        ['Europe', 210, 27],
        ['Asia Pacific', 140, 18],
    ]
    for r, row_data in enumerate(regions, 4):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 22
    ws2.column_dimensions['C'].width = 22

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
