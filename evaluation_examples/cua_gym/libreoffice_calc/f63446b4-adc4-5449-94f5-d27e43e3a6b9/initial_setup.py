"""
Initial Setup: Group rows 5-12 to create a collapsible section for Q1 detail rows
Task ID: calc_gfl_053
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_053'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "P&L"

    # Styles
    title_font = Font(name="Arial", size=14, bold=True)
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    total_font = Font(name="Arial", size=11, bold=True)
    total_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
    currency_fmt = '$#,##0'
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Row 1: Title
    ws.merge_cells("A1:J1")
    ws["A1"] = "Quarterly Profit & Loss Statement"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    # Row 2: blank

    # Row 3: Headers
    headers = ["Category", "Revenue", "COGS", "Gross Profit", "Marketing",
               "Salaries", "Rent", "Utilities", "Misc", "Net Income"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 22
    for c in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        ws.column_dimensions[c].width = 14

    # --- Q1 Detail rows (5-12) ---
    q1_details = [
        ["Product Sales - East",    58200, 23280, 34920, 4500, 18000, 3200, 850, 620, 7750],
        ["Product Sales - West",    42800, 17120, 25680, 3800, 15000, 2800, 720, 480, 2880],
        ["Service Revenue",         31500, 9450,  22050, 2200, 12000, 1600, 540, 310, 5400],
        ["Licensing Fees",          18700, 3740,  14960, 1500, 8000,  1200, 380, 220, 3660],
        ["Consulting Income",       24300, 7290,  17010, 1800, 10000, 1400, 460, 290, 3060],
        ["Subscription Revenue",    36100, 10830, 25270, 2800, 13000, 1800, 620, 350, 5700],
        ["Partner Commissions",     12400, 4960,  7440,  1100, 6000,  900,  280, 180, -1020],
        ["Training & Workshops",    8900,  2670,  6230,  800,  5000,  700,  220, 150, -640],
    ]

    # Row 4: Q1 Total
    ws.cell(row=4, column=1, value="Q1 Total").font = total_font
    ws.cell(row=4, column=1).fill = total_fill
    q1_sums = [0] * 9
    for detail in q1_details:
        for i in range(9):
            q1_sums[i] += detail[i + 1]
    for col, val in enumerate(q1_sums, 2):
        cell = ws.cell(row=4, column=col, value=val)
        cell.font = total_font
        cell.fill = total_fill
        cell.number_format = currency_fmt
        cell.border = thin_border

    # Rows 5-12: Q1 details
    for r_offset, detail in enumerate(q1_details):
        row_num = 5 + r_offset
        for col, val in enumerate(detail, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            if col >= 2:
                cell.number_format = currency_fmt
            cell.border = thin_border

    # --- Q2 Detail rows (14-21) ---
    q2_details = [
        ["Product Sales - East",    62400, 24960, 37440, 4800, 18500, 3200, 870, 640, 8930],
        ["Product Sales - West",    45900, 18360, 27540, 4100, 15500, 2800, 740, 500, 3900],
        ["Service Revenue",         34200, 10260, 23940, 2400, 12500, 1600, 560, 330, 6550],
        ["Licensing Fees",          20100, 4020,  16080, 1600, 8200,  1200, 400, 230, 4450],
        ["Consulting Income",       26800, 8040,  18760, 1900, 10500, 1400, 480, 300, 4180],
        ["Subscription Revenue",    39500, 11850, 27650, 3000, 13500, 1800, 640, 370, 8340],
        ["Partner Commissions",     14100, 5640,  8460,  1200, 6200,  900,  290, 190, -320],
        ["Training & Workshops",    10200, 3060,  7140,  900,  5200,  700,  230, 160, 950],
    ]

    # Row 13: Q2 Total
    ws.cell(row=13, column=1, value="Q2 Total").font = total_font
    ws.cell(row=13, column=1).fill = total_fill
    q2_sums = [0] * 9
    for detail in q2_details:
        for i in range(9):
            q2_sums[i] += detail[i + 1]
    for col, val in enumerate(q2_sums, 2):
        cell = ws.cell(row=13, column=col, value=val)
        cell.font = total_font
        cell.fill = total_fill
        cell.number_format = currency_fmt
        cell.border = thin_border

    # Rows 14-21: Q2 details
    for r_offset, detail in enumerate(q2_details):
        row_num = 14 + r_offset
        for col, val in enumerate(detail, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            if col >= 2:
                cell.number_format = currency_fmt
            cell.border = thin_border

    # --- Q3 Detail rows (23-30) ---
    q3_details = [
        ["Product Sales - East",    55800, 22320, 33480, 4300, 18200, 3300, 890, 650, 6140],
        ["Product Sales - West",    41200, 16480, 24720, 3600, 15200, 2900, 750, 510, 1760],
        ["Service Revenue",         29800, 8940,  20860, 2100, 11800, 1650, 550, 320, 4440],
        ["Licensing Fees",          17500, 3500,  14000, 1400, 7800,  1250, 390, 230, 2930],
        ["Consulting Income",       22900, 6870,  16030, 1700, 9800,  1450, 470, 300, 2310],
        ["Subscription Revenue",    34600, 10380, 24220, 2600, 12800, 1850, 630, 360, 5980],
        ["Partner Commissions",     11600, 4640,  6960,  1000, 5900,  950,  290, 190, -1370],
        ["Training & Workshops",    8200,  2460,  5740,  750,  4800,  720,  230, 160, -920],
    ]

    # Row 22: Q3 Total
    ws.cell(row=22, column=1, value="Q3 Total").font = total_font
    ws.cell(row=22, column=1).fill = total_fill
    q3_sums = [0] * 9
    for detail in q3_details:
        for i in range(9):
            q3_sums[i] += detail[i + 1]
    for col, val in enumerate(q3_sums, 2):
        cell = ws.cell(row=22, column=col, value=val)
        cell.font = total_font
        cell.fill = total_fill
        cell.number_format = currency_fmt
        cell.border = thin_border

    # Rows 23-30: Q3 details
    for r_offset, detail in enumerate(q3_details):
        row_num = 23 + r_offset
        for col, val in enumerate(detail, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            if col >= 2:
                cell.number_format = currency_fmt
            cell.border = thin_border

    # NO row grouping in initial state - that is the task
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
