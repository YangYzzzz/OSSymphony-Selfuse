"""
Initial Setup: Count 'Failed' status entries using COUNTIF formula
Task ID: calc_fmb_countif_text_009
Domain: libreoffice_calc
"""

import random
from datetime import date, timedelta
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_countif_text_009'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'QC Log'

    # --- Row 1: Headers ---
    headers = ['Batch ID', 'Product', 'Inspector', 'Date', 'Status', 'Notes']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic product names
    products = [
        'Circuit Board A100', 'Valve Assembly V-30', 'Pressure Sensor P-12',
        'Hydraulic Pump HP-5', 'Steel Rod SR-20', 'Gasket Kit GK-8',
        'Bearing Housing BH-4', 'Thermal Sensor TS-9', 'Actuator Module AM-3',
        'Relay Switch RS-7', 'Control Panel CP-15', 'Filter Cartridge FC-11',
        'Motor Coupling MC-6', 'Gear Assembly GA-22', 'Safety Valve SV-14',
    ]

    # Realistic inspector names
    inspectors = [
        'Sarah Chen', 'Marcus Johnson', 'Priya Nair', 'David Kim',
        'Emma Russo', 'James Okafor', 'Linda Tran', 'Omar Al-Farsi',
        'Helen Brooks', 'Carlos Mendez',
    ]

    # Notes templates
    pass_notes = [
        'All parameters within spec', 'Passed visual inspection',
        'Dimensions confirmed acceptable', 'No defects detected',
        'Surface finish meets standard', 'Tolerances verified OK',
        'Functional test passed', 'Batch certified clean',
        'Quality confirmed satisfactory', 'Inspection complete - pass',
    ]
    fail_notes = [
        'Surface crack detected on unit 3', 'Dimension out of tolerance by 0.5mm',
        'Coating irregularity found', 'Weld seam defect observed',
        'Pressure test failed at 150 PSI', 'Thread damage on connector port',
        'Calibration drift detected', 'Contamination found in batch',
        'Warping beyond acceptable limit', 'Electrical continuity failed',
    ]
    pending_notes = [
        'Awaiting senior review', 'Additional testing required',
        'Referred to engineering for assessment', 'Pending material certification',
        'Awaiting customer specification update', 'On hold - further analysis needed',
        'Second opinion requested', 'Documentation pending from supplier',
    ]

    # Build status list: exactly 43 Failed, 187 Passed, 20 Pending Review
    statuses = (['Failed'] * 43) + (['Passed'] * 187) + (['Pending Review'] * 20)
    random.seed(42)
    random.shuffle(statuses)

    start_date = date(2024, 1, 2)

    for i, status in enumerate(statuses):
        row = i + 2
        batch_num = 1000 + i
        batch_id = f'BT-2024-{batch_num:04d}'
        product = products[i % len(products)]
        inspector = inspectors[i % len(inspectors)]
        rec_date = start_date + timedelta(days=i // 5)

        if status == 'Failed':
            note = fail_notes[i % len(fail_notes)]
        elif status == 'Pending Review':
            note = pending_notes[i % len(pending_notes)]
        else:
            note = pass_notes[i % len(pass_notes)]

        ws.cell(row=row, column=1, value=batch_id)
        ws.cell(row=row, column=2, value=product)
        ws.cell(row=row, column=3, value=inspector)
        ws.cell(row=row, column=4, value=rec_date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=5, value=status)
        ws.cell(row=row, column=6, value=note)

    # G2 label as specified in context
    ws.cell(row=2, column=7, value='Failed Count')
    # H2 is intentionally EMPTY (target cell for the COUNTIF formula)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: QC Log')
    print(f'  Rows: 250 data rows (rows 2-251)')
    print(f'  Statuses: 43 Failed, 187 Passed, 20 Pending Review')
    print(f'  G2: "Failed Count" label')
    print(f'  H2: empty (target for COUNTIF formula)')

create_initial()
