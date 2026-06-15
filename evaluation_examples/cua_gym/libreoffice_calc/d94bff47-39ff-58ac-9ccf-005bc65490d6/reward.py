"""
Reward Script: Count 'Failed' status entries using COUNTIF formula in H2
Task ID: calc_fmb_countif_text_009
Domain: libreoffice_calc
Scoring:
  Component 1 (0.6 pts): Cell H2 contains a COUNTIF formula with 'Failed' criteria
  Component 2 (0.4 pts): The COUNTIF formula references the correct range E2:E251
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmb_countif_text_009'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Task: Place =COUNTIF(E2:E251,"Failed") in cell H2 of the 'QC Log' sheet.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition gate: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: 'QC Log' sheet must exist
    if 'QC Log' not in wb.sheetnames:
        print("CRITICAL: Sheet 'QC Log' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['QC Log']

    # Component 1: H2 contains a COUNTIF formula that targets 'Failed' (0.6 points)
    # This FAILS on initial (H2 is empty/None) and PASSES on golden (has COUNTIF formula)
    try:
        h2_value = ws['H2'].value
        # Check that it is a formula string containing COUNTIF and "Failed"
        if isinstance(h2_value, str):
            h2_upper = h2_value.upper().replace(' ', '')
            has_countif = 'COUNTIF' in h2_upper
            has_failed = '"FAILED"' in h2_upper or "'FAILED'" in h2_upper
            if has_countif and has_failed:
                print(f"PASS: Component 1 — H2 contains COUNTIF formula with 'Failed' criteria (value: {h2_value}) (0.6 pts)")
                total_score += 0.6
            elif has_countif:
                print(f"FAIL: Component 1 — H2 contains COUNTIF but 'Failed' criteria not found (value: {h2_value})")
            else:
                print(f"FAIL: Component 1 — H2 does not contain a COUNTIF formula (value: {repr(h2_value)})")
        else:
            print(f"FAIL: Component 1 — H2 is not a formula string (value: {repr(h2_value)})")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not check H2: {e}")

    # Component 2: The COUNTIF formula references the correct range E2:E251 (0.4 points)
    # This FAILS on initial (H2 is empty) and PASSES on golden (has correct range)
    try:
        h2_value = ws['H2'].value
        if isinstance(h2_value, str):
            h2_upper = h2_value.upper().replace(' ', '')
            has_correct_range = 'E2:E251' in h2_upper
            if has_correct_range:
                print(f"PASS: Component 2 — COUNTIF formula uses correct range E2:E251 (0.4 pts)")
                total_score += 0.4
            else:
                print(f"FAIL: Component 2 — Expected range E2:E251 in formula, found: {h2_value}")
        else:
            print(f"FAIL: Component 2 — H2 is not a formula string, cannot check range (value: {repr(h2_value)})")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not check range in H2: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
