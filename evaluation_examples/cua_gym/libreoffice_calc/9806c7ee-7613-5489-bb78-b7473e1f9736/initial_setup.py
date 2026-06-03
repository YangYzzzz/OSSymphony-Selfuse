"""
Initial Setup: Hong Kong Dim Sum Restaurant Lookup Task
Task ID: osworld_multi_apps_restaurant_lookup_011
Domain: libreoffice_calc (multi-app: gedit + LibreOffice Calc)

Creates:
  - /home/user/Desktop/hk_dimsum.txt  — list of 6 HK dim sum restaurant names
  - /home/user/Desktop/DIM_SUM.xlsx   — spreadsheet with Name col pre-filled, Address/Website/Phone empty

GUI: Opens hk_dimsum.txt in gedit and DIM_SUM.xlsx in LibreOffice Calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

WORKDIR = '/home/user/Desktop'
TASK_ID = 'osworld_multi_apps_restaurant_lookup_011'
XLSX_OUTPUT = f'{WORKDIR}/DIM_SUM.xlsx'
TXT_OUTPUT = f'{WORKDIR}/hk_dimsum.txt'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_txt():
    """Create the hk_dimsum.txt file with 6 restaurant names."""
    content = (
        "Hong Kong Dim Sum Restaurants\n"
        "==============================\n"
        "1. Tim Ho Wan (Sham Shui Po)\n"
        "2. One Dim Sum\n"
        "3. Lin Heung Tea House\n"
        "4. Maxim's Palace\n"
        "5. Crystal Jade\n"
        "6. The Chairman\n"
    )
    os.makedirs(WORKDIR, exist_ok=True)
    with open(TXT_OUTPUT, 'w', encoding='utf-8') as f:
        f.write(content)
    print(f'Text file created: {TXT_OUTPUT}')


def create_xlsx():
    """Create DIM_SUM.xlsx with restaurant names pre-filled but Address/Website/Phone empty."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Dim Sum Restaurants'

    # Header row with bold formatting
    headers = ['Name', 'Address', 'Website', 'Phone']
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Restaurant names pre-filled in column A (rows 1-6 means data rows, with header in row 1)
    # Context says "Restaurant names pre-filled in column A rows 1-6" — interpreted as data rows 2-7
    restaurants = [
        'Tim Ho Wan (Sham Shui Po)',
        'One Dim Sum',
        'Lin Heung Tea House',
        "Maxim's Palace",
        'Crystal Jade',
        'The Chairman',
    ]

    for row_idx, name in enumerate(restaurants, 2):
        ws.cell(row=row_idx, column=1, value=name)
        # Columns B (Address), C (Website), D (Phone) are intentionally LEFT EMPTY

    # Set column widths for readability
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 32
    ws.column_dimensions['D'].width = 20

    # Row height for header
    ws.row_dimensions[1].height = 20

    wb.save(XLSX_OUTPUT)
    print(f'Spreadsheet created: {XLSX_OUTPUT}')


def main():
    create_txt()
    create_xlsx()

    # GUI-ready startup: open hk_dimsum.txt in gedit first
    launch_gui(f'gedit "{TXT_OUTPUT}"', delay_sec=2.0)

    # Then open DIM_SUM.xlsx in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{XLSX_OUTPUT}"', delay_sec=2.0)

    print('GUI_READY: Launched gedit with hk_dimsum.txt and LibreOffice Calc with DIM_SUM.xlsx (DISPLAY=:0)')


main()
