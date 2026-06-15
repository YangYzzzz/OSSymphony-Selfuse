"""
Reward Script: Fill in dim sum restaurant details (Address, Website, Phone) in DIM_SUM.xlsx
Task ID: osworld_multi_apps_restaurant_lookup_011
Domain: libreoffice_calc
Scoring:
  Component 1: Tim Ho Wan row filled with correct phone anchor (0.20 pts)
  Component 2: One Dim Sum row filled (0.16 pts)
  Component 3: Lin Heung Tea House row filled (0.16 pts)
  Component 4: Maxim's Palace row filled (0.16 pts)
  Component 5: Crystal Jade row filled (0.16 pts)
  Component 6: The Chairman row filled (0.16 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user/Desktop'
TASK_ID = 'osworld_multi_apps_restaurant_lookup_011'
FILE_PATH = f'{WORKDIR}/DIM_SUM.xlsx'

# Expected restaurant name in each row (for context / diagnostics)
RESTAURANT_ROWS = {
    2: 'Tim Ho Wan (Sham Shui Po)',
    3: 'One Dim Sum',
    4: 'Lin Heung Tea House',
    5: "Maxim's Palace",
    6: 'Crystal Jade',
    7: 'The Chairman',
}

# Ground truth phone for Tim Ho Wan — used to anchor data accuracy
TIM_HO_WAN_PHONE = '+852 2788 1226'


def all_fields_filled(ws, row):
    """Return True if Address (col B), Website (col C), and Phone (col D) are all non-empty."""
    address = ws.cell(row=row, column=2).value
    website = ws.cell(row=row, column=3).value
    phone   = ws.cell(row=row, column=4).value
    return (
        address is not None and str(address).strip() != '' and
        website is not None and str(website).strip() != '' and
        phone   is not None and str(phone).strip()   != ''
    )


def verify_task(file_path):
    """
    Verify that all 6 dim sum restaurants have their Address, Website, and Phone filled in.
    Uses the Tim Ho Wan phone number as a ground-truth anchor.
    Returns a float between 0.0 and 1.0.
    """
    total_score = 0.0

    # Precondition: file must load
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Find the active / only worksheet
    try:
        ws = wb.active
        print(f"INFO: Loaded sheet '{ws.title}', dimensions {ws.dimensions}")
    except Exception as e:
        print(f"CRITICAL: Cannot access worksheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: verify header row integrity (gate — not scored)
    try:
        header_name    = ws.cell(row=1, column=1).value
        header_address = ws.cell(row=1, column=2).value
        header_website = ws.cell(row=1, column=3).value
        header_phone   = ws.cell(row=1, column=4).value
        if header_name != 'Name':
            print(f"WARN: Header A1 is {repr(header_name)}, expected 'Name'. Proceeding anyway.")
    except Exception as e:
        print(f"WARN: Could not read headers: {e}. Proceeding anyway.")

    # -------------------------------------------------------------------------
    # Component 1: Tim Ho Wan (row 2) — all 3 fields filled, phone anchor (0.20 pts)
    # -------------------------------------------------------------------------
    try:
        address = ws.cell(row=2, column=2).value
        website = ws.cell(row=2, column=3).value
        phone   = ws.cell(row=2, column=4).value

        fields_filled = (
            address is not None and str(address).strip() != '' and
            website is not None and str(website).strip() != '' and
            phone   is not None and str(phone).strip()   != ''
        )
        phone_correct = phone is not None and str(phone).strip() == TIM_HO_WAN_PHONE

        if fields_filled and phone_correct:
            print(f"PASS: Component 1 — Tim Ho Wan: Address='{address}', Website='{website}', Phone='{phone}' (0.20 pts)")
            total_score += 0.20
        elif fields_filled:
            # Partial: fields filled but phone doesn't exactly match ground truth
            # Still award partial credit (0.10) since data may have minor formatting difference
            print(f"PARTIAL: Component 1 — Tim Ho Wan fields filled but phone '{phone}' != expected '{TIM_HO_WAN_PHONE}' (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 — Tim Ho Wan: Address={repr(address)}, Website={repr(website)}, Phone={repr(phone)}")
    except Exception as e:
        print(f"ERROR: Component 1 — Tim Ho Wan: {e}")

    # -------------------------------------------------------------------------
    # Component 2: One Dim Sum (row 3) — all 3 fields filled (0.16 pts)
    # -------------------------------------------------------------------------
    try:
        if all_fields_filled(ws, 3):
            address = ws.cell(row=3, column=2).value
            phone   = ws.cell(row=3, column=4).value
            print(f"PASS: Component 2 — One Dim Sum filled: Address='{address}', Phone='{phone}' (0.16 pts)")
            total_score += 0.16
        else:
            address = ws.cell(row=3, column=2).value
            website = ws.cell(row=3, column=3).value
            phone   = ws.cell(row=3, column=4).value
            print(f"FAIL: Component 2 — One Dim Sum: Address={repr(address)}, Website={repr(website)}, Phone={repr(phone)}")
    except Exception as e:
        print(f"ERROR: Component 2 — One Dim Sum: {e}")

    # -------------------------------------------------------------------------
    # Component 3: Lin Heung Tea House (row 4) — all 3 fields filled (0.16 pts)
    # -------------------------------------------------------------------------
    try:
        if all_fields_filled(ws, 4):
            address = ws.cell(row=4, column=2).value
            phone   = ws.cell(row=4, column=4).value
            print(f"PASS: Component 3 — Lin Heung Tea House filled: Address='{address}', Phone='{phone}' (0.16 pts)")
            total_score += 0.16
        else:
            address = ws.cell(row=4, column=2).value
            website = ws.cell(row=4, column=3).value
            phone   = ws.cell(row=4, column=4).value
            print(f"FAIL: Component 3 — Lin Heung Tea House: Address={repr(address)}, Website={repr(website)}, Phone={repr(phone)}")
    except Exception as e:
        print(f"ERROR: Component 3 — Lin Heung Tea House: {e}")

    # -------------------------------------------------------------------------
    # Component 4: Maxim's Palace (row 5) — all 3 fields filled (0.16 pts)
    # -------------------------------------------------------------------------
    try:
        if all_fields_filled(ws, 5):
            address = ws.cell(row=5, column=2).value
            phone   = ws.cell(row=5, column=4).value
            print(f"PASS: Component 4 — Maxim's Palace filled: Address='{address}', Phone='{phone}' (0.16 pts)")
            total_score += 0.16
        else:
            address = ws.cell(row=5, column=2).value
            website = ws.cell(row=5, column=3).value
            phone   = ws.cell(row=5, column=4).value
            print(f"FAIL: Component 4 — Maxim's Palace: Address={repr(address)}, Website={repr(website)}, Phone={repr(phone)}")
    except Exception as e:
        print(f"ERROR: Component 4 — Maxim's Palace: {e}")

    # -------------------------------------------------------------------------
    # Component 5: Crystal Jade (row 6) — all 3 fields filled (0.16 pts)
    # -------------------------------------------------------------------------
    try:
        if all_fields_filled(ws, 6):
            address = ws.cell(row=6, column=2).value
            phone   = ws.cell(row=6, column=4).value
            print(f"PASS: Component 5 — Crystal Jade filled: Address='{address}', Phone='{phone}' (0.16 pts)")
            total_score += 0.16
        else:
            address = ws.cell(row=6, column=2).value
            website = ws.cell(row=6, column=3).value
            phone   = ws.cell(row=6, column=4).value
            print(f"FAIL: Component 5 — Crystal Jade: Address={repr(address)}, Website={repr(website)}, Phone={repr(phone)}")
    except Exception as e:
        print(f"ERROR: Component 5 — Crystal Jade: {e}")

    # -------------------------------------------------------------------------
    # Component 6: The Chairman (row 7) — all 3 fields filled (0.16 pts)
    # -------------------------------------------------------------------------
    try:
        if all_fields_filled(ws, 7):
            address = ws.cell(row=7, column=2).value
            phone   = ws.cell(row=7, column=4).value
            print(f"PASS: Component 6 — The Chairman filled: Address='{address}', Phone='{phone}' (0.16 pts)")
            total_score += 0.16
        else:
            address = ws.cell(row=7, column=2).value
            website = ws.cell(row=7, column=3).value
            phone   = ws.cell(row=7, column=4).value
            print(f"FAIL: Component 6 — The Chairman: Address={repr(address)}, Website={repr(website)}, Phone={repr(phone)}")
    except Exception as e:
        print(f"ERROR: Component 6 — The Chairman: {e}")

    # -------------------------------------------------------------------------
    # Final score
    # -------------------------------------------------------------------------
    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score:.4f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point: reward scripts always run on the VM
if not os.path.exists(FILE_PATH):
    print(f"File not found: {FILE_PATH}")
    print("REWARD: 0.0")
else:
    verify_task(FILE_PATH)
