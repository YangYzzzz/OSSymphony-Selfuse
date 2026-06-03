"""
Initial Setup: Lab measurement spreadsheet for decimal places formatting task
Task ID: calc_fmt_numfmt_decimal_places_021
Domain: libreoffice_calc

Creates a spreadsheet with one sheet 'Lab Results' containing 29 rows of
realistic lab measurement data. Column B uses 'General' format (no fixed
decimal places) intentionally — the task requires setting it to 3 decimal places.
"""

import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_numfmt_decimal_places_021'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Lab Results'

    # --- Headers (row 1) ---
    headers = ['Sample ID', 'Measurement (mm)', 'Tolerance', 'Pass/Fail']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # --- Column widths ---
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12

    # --- Data rows 2-30 (29 samples) ---
    # Realistic lab measurement data with varying decimal places in 'General' format
    data = [
        ('SMP-001', 12.4,        0.05, 'Pass'),
        ('SMP-002', 8.9,         0.05, 'Pass'),
        ('SMP-003', 15.234567,   0.05, 'Fail'),
        ('SMP-004', 10.75,       0.05, 'Pass'),
        ('SMP-005', 7.3,         0.05, 'Pass'),
        ('SMP-006', 14.001,      0.05, 'Pass'),
        ('SMP-007', 9.85,        0.05, 'Pass'),
        ('SMP-008', 11.6,        0.05, 'Pass'),
        ('SMP-009', 13.489,      0.05, 'Fail'),
        ('SMP-010', 6.22,        0.05, 'Pass'),
        ('SMP-011', 10.0,        0.05, 'Pass'),
        ('SMP-012', 12.8,        0.05, 'Pass'),
        ('SMP-013', 5.567,       0.05, 'Fail'),
        ('SMP-014', 9.1,         0.05, 'Pass'),
        ('SMP-015', 11.45,       0.05, 'Pass'),
        ('SMP-016', 14.72,       0.05, 'Pass'),
        ('SMP-017', 8.003,       0.05, 'Pass'),
        ('SMP-018', 13.6,        0.05, 'Pass'),
        ('SMP-019', 7.89,        0.05, 'Pass'),
        ('SMP-020', 10.555,      0.05, 'Fail'),
        ('SMP-021', 12.1,        0.05, 'Pass'),
        ('SMP-022', 9.34,        0.05, 'Pass'),
        ('SMP-023', 6.7,         0.05, 'Pass'),
        ('SMP-024', 11.92,       0.05, 'Pass'),
        ('SMP-025', 15.04,       0.05, 'Fail'),
        ('SMP-026', 8.5,         0.05, 'Pass'),
        ('SMP-027', 13.21,       0.05, 'Pass'),
        ('SMP-028', 10.88,       0.05, 'Pass'),
        ('SMP-029', 7.643,       0.05, 'Pass'),
    ]

    for r, (sample_id, measurement, tolerance, pass_fail) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=sample_id)
        # Column B: raw float, 'General' format (no number_format set)
        ws.cell(row=r, column=2, value=measurement)
        ws.cell(row=r, column=3, value=tolerance)
        ws.cell(row=r, column=4, value=pass_fail)

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Lab Results')
    print(f'  Data rows: 2-30 (29 samples)')
    print(f'  Column B number_format: General (no fixed decimal places)')


create_initial()
