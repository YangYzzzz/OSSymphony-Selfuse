"""
Reward Script: Set measurement values in cells B2:B30 to display exactly 3 decimal places.
Task ID: calc_fmt_numfmt_decimal_places_021
Domain: libreoffice_calc
Scoring:
  Component 1: Cells B2:B15 (first half) all have number format '0.000'         (0.5 pts)
               AND underlying values in those cells are unchanged
  Component 2: Cells B16:B30 (second half) all have number format '0.000'       (0.3 pts)
               AND underlying values in those cells are unchanged
  Component 3: Header B1 format unchanged + no other columns modified             (0.2 pts)
               (only passes when format '0.000' changes exist while non-task cells are clean)
  Total: 1.0

NOTE: Components 1 and 2 intentionally split the range to allow partial credit.
      Components are anchored to the task change (format change) so they FAIL on initial.
      Component 3 tests the 'no collateral damage' condition by verifying non-B2:B30 cells
      did NOT receive the '0.000' format (which would be false on initial where B1 is General).
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_numfmt_decimal_places_021'

# Expected underlying values in B2:B30 (must remain unchanged after formatting)
EXPECTED_B_VALUES = [
    12.4, 8.9, 15.234567, 10.75, 7.3, 14.001, 9.85, 11.6, 13.489,
    6.22, 10, 12.8, 5.567, 9.1, 11.45, 14.72, 8.003, 13.6, 7.89,
    10.555, 12.1, 9.34, 6.7, 11.92, 15.04, 8.5, 13.21, 10.88, 7.643
]


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: check sheet 'Lab Results' exists
    if 'Lab Results' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Lab Results' not found — file structure broken.")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Lab Results']

    # Component 1: Cells B2:B15 have number format '0.000' AND values unchanged (0.5 points)
    # This FAILS on the initial file (format is 'General' there) — tests the actual task change.
    # Sub-condition: underlying values must be preserved (compound check for data integrity).
    try:
        format_ok_count = 0
        value_ok_count = 0
        comp1_issues = []

        for i, row in enumerate(range(2, 16)):  # rows 2-15 = first 14 cells
            cell = ws.cell(row=row, column=2)
            fmt = cell.number_format
            actual_val = cell.value
            expected_val = EXPECTED_B_VALUES[i]

            fmt_ok = (fmt == '0.000')
            try:
                val_ok = (actual_val is not None and abs(float(actual_val) - expected_val) < 1e-9)
            except (TypeError, ValueError):
                val_ok = False

            if fmt_ok:
                format_ok_count += 1
            if val_ok:
                value_ok_count += 1

            if not fmt_ok or not val_ok:
                comp1_issues.append(
                    f"B{row}: fmt={repr(fmt)}, val={actual_val}(expected {expected_val})"
                )

        # Both format and value must be correct for full points
        if format_ok_count == 14 and value_ok_count == 14:
            print(f"PASS: Component 1 — All B2:B15 have '0.000' format with unchanged values (0.5 pts)")
            total_score += 0.5
        elif format_ok_count > 0:
            partial = round(0.5 * format_ok_count / 14, 4)
            total_score += partial
            print(f"PARTIAL: Component 1 — {format_ok_count}/14 cells in B2:B15 have '0.000' format "
                  f"({partial} pts); issues: {comp1_issues[:3]}")
        else:
            print(f"FAIL: Component 1 — No cells in B2:B15 have '0.000' format. "
                  f"Issues: {comp1_issues[:3]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Cells B16:B30 have number format '0.000' AND values unchanged (0.3 points)
    # This FAILS on the initial file (format is 'General' there) — tests the actual task change.
    try:
        format_ok_count = 0
        value_ok_count = 0
        comp2_issues = []

        for i, row in enumerate(range(16, 31)):  # rows 16-30 = last 15 cells
            cell = ws.cell(row=row, column=2)
            fmt = cell.number_format
            actual_val = cell.value
            expected_val = EXPECTED_B_VALUES[14 + i]  # offset by 14 (B2:B15 is first 14)

            fmt_ok = (fmt == '0.000')
            try:
                val_ok = (actual_val is not None and abs(float(actual_val) - expected_val) < 1e-9)
            except (TypeError, ValueError):
                val_ok = False

            if fmt_ok:
                format_ok_count += 1
            if val_ok:
                value_ok_count += 1

            if not fmt_ok or not val_ok:
                comp2_issues.append(
                    f"B{row}: fmt={repr(fmt)}, val={actual_val}(expected {expected_val})"
                )

        if format_ok_count == 15 and value_ok_count == 15:
            print(f"PASS: Component 2 — All B16:B30 have '0.000' format with unchanged values (0.3 pts)")
            total_score += 0.3
        elif format_ok_count > 0:
            partial = round(0.3 * format_ok_count / 15, 4)
            total_score += partial
            print(f"PARTIAL: Component 2 — {format_ok_count}/15 cells in B16:B30 have '0.000' format "
                  f"({partial} pts); issues: {comp2_issues[:3]}")
        else:
            print(f"FAIL: Component 2 — No cells in B16:B30 have '0.000' format. "
                  f"Issues: {comp2_issues[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: No collateral damage — B1 header NOT given '0.000' format,
    # and other columns (A, C, D) NOT given '0.000' format (0.2 points)
    # This PASSES on the initial file where everything is 'General',
    # but the check is ANCHORED to detecting if the task was done too broadly (over-formatting).
    # We verify: B1 does NOT have '0.000' format, AND total '0.000' formatted cells == exactly 29.
    # This FAILS on initial because count of '0.000' cells = 0, not 29.
    # This FAILS on a bad golden where header was also formatted.
    try:
        b1_fmt = ws.cell(row=1, column=2).number_format
        b1_not_overformatted = (b1_fmt != '0.000')

        # Count total cells in sheet that have '0.000' format (should be exactly 29)
        total_000_format_count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.number_format == '0.000':
                    total_000_format_count += 1

        # Exactly 29 cells (B2:B30) should have the format — not more, not fewer
        exact_count_ok = (total_000_format_count == 29)

        if b1_not_overformatted and exact_count_ok:
            print(f"PASS: Component 3 — Exactly 29 cells have '0.000' format (B2:B30 only), "
                  f"B1 and other columns unaffected (0.2 pts)")
            total_score += 0.2
        else:
            issues = []
            if not b1_not_overformatted:
                issues.append(f"B1 incorrectly has '0.000' format")
            if not exact_count_ok:
                issues.append(f"Expected exactly 29 cells with '0.000', found {total_000_format_count}")
            print(f"FAIL: Component 3 — Collateral damage detected: {issues}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
