"""
Initial Setup: Expense log spreadsheet with date column but no validation
Task ID: calc_gao_007
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from datetime import date, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_gao_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"

    # Headers
    headers = ["ID", "Description", "Amount", "Date", "Category"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic expense data in rows 2-15
    today = date.today()
    data = [
        ["EXP-001", "Office supplies - printer paper", 45.99, today - timedelta(days=3), "Office Supplies"],
        ["EXP-002", "Client lunch meeting at Riverside Grill", 127.50, today - timedelta(days=7), "Meals & Entertainment"],
        ["EXP-003", "Monthly software subscription - Figma", 15.00, today - timedelta(days=14), "Software"],
        ["EXP-004", "Taxi to airport for Chicago conference", 38.75, today - timedelta(days=21), "Travel"],
        ["EXP-005", "Hotel stay - 2 nights in Chicago", 389.00, today - timedelta(days=20), "Travel"],
        ["EXP-006", "Conference registration fee", 250.00, today - timedelta(days=30), "Training"],
        ["EXP-007", "New ergonomic keyboard - Logitech MX Keys", 99.99, today - timedelta(days=45), "Equipment"],
        ["EXP-008", "Team building dinner at Sakura Japanese", 215.80, today - timedelta(days=60), "Meals & Entertainment"],
        ["EXP-009", "Parking garage monthly pass - March", 85.00, today - timedelta(days=75), "Transportation"],
        ["EXP-010", "Cloud hosting fees - AWS Q1", 342.17, today - timedelta(days=90), "Software"],
        ["EXP-011", "Business cards reprint - 500 qty", 62.00, today - timedelta(days=120), "Office Supplies"],
        ["EXP-012", "Flight to Seattle for partner meeting", 478.50, today - timedelta(days=150), "Travel"],
        ["EXP-013", "External monitor - Dell UltraSharp 27", 449.99, today - timedelta(days=200), "Equipment"],
        ["EXP-014", "Annual domain renewal - companysite.com", 14.99, today - timedelta(days=250), "Software"],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Format date column
    for r in range(2, 16):
        ws.cell(row=r, column=4).number_format = 'yyyy-mm-dd'

    # Set reasonable column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 22

    # Format Amount as currency
    for r in range(2, 16):
        ws.cell(row=r, column=3).number_format = '$#,##0.00'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
