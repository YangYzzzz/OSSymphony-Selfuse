"""
Initial Setup: Create workbook with 5 monthly sheets (Jan-May), all unprotected, no macros.
Task ID: calc_mcp_017
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_017'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    wb = openpyxl.Workbook()

    # Common styles
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    currency_fmt = '$#,##0.00'
    date_fmt = 'yyyy-mm-dd'
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # Monthly data templates - realistic sales data for a retail company
    months_data = {
        'Jan': {
            'headers': ['Date', 'Sales Rep', 'Region', 'Product', 'Units Sold', 'Unit Price', 'Total Revenue'],
            'data': [
                ['2025-01-03', 'Sarah Chen', 'Northeast', 'Widget Pro X1', 45, 129.99, 5849.55],
                ['2025-01-05', 'Marcus Johnson', 'Southeast', 'Gadget Elite S2', 32, 249.50, 7984.00],
                ['2025-01-08', 'Emily Rodriguez', 'West Coast', 'Widget Pro X1', 58, 129.99, 7539.42],
                ['2025-01-10', 'David Park', 'Midwest', 'Component Alpha 3', 120, 45.75, 5490.00],
                ['2025-01-12', 'Sarah Chen', 'Northeast', 'Gadget Elite S2', 27, 249.50, 6736.50],
                ['2025-01-15', 'Rachel Kim', 'West Coast', 'Widget Pro X1', 63, 129.99, 8189.37],
                ['2025-01-18', 'James Wilson', 'Southeast', 'Component Alpha 3', 95, 45.75, 4346.25],
                ['2025-01-20', 'Marcus Johnson', 'Southeast', 'Premium Bundle Z', 15, 599.00, 8985.00],
                ['2025-01-23', 'Emily Rodriguez', 'West Coast', 'Gadget Elite S2', 41, 249.50, 10229.50],
                ['2025-01-25', 'David Park', 'Midwest', 'Widget Pro X1', 52, 129.99, 6759.48],
                ['2025-01-28', 'Rachel Kim', 'West Coast', 'Premium Bundle Z', 18, 599.00, 10782.00],
                ['2025-01-30', 'James Wilson', 'Southeast', 'Component Alpha 3', 110, 45.75, 5032.50],
            ],
        },
        'Feb': {
            'headers': ['Date', 'Sales Rep', 'Region', 'Product', 'Units Sold', 'Unit Price', 'Total Revenue'],
            'data': [
                ['2025-02-02', 'Sarah Chen', 'Northeast', 'Gadget Elite S2', 38, 249.50, 9481.00],
                ['2025-02-04', 'Marcus Johnson', 'Southeast', 'Widget Pro X1', 55, 129.99, 7149.45],
                ['2025-02-07', 'Emily Rodriguez', 'West Coast', 'Premium Bundle Z', 22, 599.00, 13178.00],
                ['2025-02-10', 'David Park', 'Midwest', 'Component Alpha 3', 88, 45.75, 4026.00],
                ['2025-02-12', 'Rachel Kim', 'West Coast', 'Widget Pro X1', 67, 129.99, 8709.33],
                ['2025-02-14', 'James Wilson', 'Southeast', 'Gadget Elite S2', 29, 249.50, 7235.50],
                ['2025-02-17', 'Sarah Chen', 'Northeast', 'Component Alpha 3', 135, 45.75, 6176.25],
                ['2025-02-19', 'Marcus Johnson', 'Southeast', 'Premium Bundle Z', 12, 599.00, 7188.00],
                ['2025-02-22', 'Emily Rodriguez', 'West Coast', 'Widget Pro X1', 48, 129.99, 6239.52],
                ['2025-02-25', 'David Park', 'Midwest', 'Gadget Elite S2', 35, 249.50, 8732.50],
                ['2025-02-27', 'Rachel Kim', 'West Coast', 'Component Alpha 3', 102, 45.75, 4666.50],
            ],
        },
        'Mar': {
            'headers': ['Date', 'Sales Rep', 'Region', 'Product', 'Units Sold', 'Unit Price', 'Total Revenue'],
            'data': [
                ['2025-03-01', 'James Wilson', 'Southeast', 'Widget Pro X1', 61, 129.99, 7929.39],
                ['2025-03-04', 'Sarah Chen', 'Northeast', 'Premium Bundle Z', 25, 599.00, 14975.00],
                ['2025-03-06', 'Marcus Johnson', 'Southeast', 'Gadget Elite S2', 44, 249.50, 10978.00],
                ['2025-03-09', 'Emily Rodriguez', 'West Coast', 'Component Alpha 3', 150, 45.75, 6862.50],
                ['2025-03-12', 'David Park', 'Midwest', 'Widget Pro X1', 39, 129.99, 5069.61],
                ['2025-03-14', 'Rachel Kim', 'West Coast', 'Gadget Elite S2', 51, 249.50, 12724.50],
                ['2025-03-17', 'James Wilson', 'Southeast', 'Premium Bundle Z', 19, 599.00, 11381.00],
                ['2025-03-20', 'Sarah Chen', 'Northeast', 'Component Alpha 3', 78, 45.75, 3568.50],
                ['2025-03-23', 'Marcus Johnson', 'Southeast', 'Widget Pro X1', 72, 129.99, 9359.28],
                ['2025-03-26', 'Emily Rodriguez', 'West Coast', 'Premium Bundle Z', 14, 599.00, 8386.00],
                ['2025-03-28', 'David Park', 'Midwest', 'Gadget Elite S2', 33, 249.50, 8233.50],
                ['2025-03-31', 'Rachel Kim', 'West Coast', 'Widget Pro X1', 56, 129.99, 7279.44],
            ],
        },
        'Apr': {
            'headers': ['Date', 'Sales Rep', 'Region', 'Product', 'Units Sold', 'Unit Price', 'Total Revenue'],
            'data': [
                ['2025-04-02', 'James Wilson', 'Southeast', 'Component Alpha 3', 125, 45.75, 5718.75],
                ['2025-04-04', 'Sarah Chen', 'Northeast', 'Widget Pro X1', 49, 129.99, 6369.51],
                ['2025-04-07', 'Marcus Johnson', 'Southeast', 'Premium Bundle Z', 17, 599.00, 10183.00],
                ['2025-04-10', 'Emily Rodriguez', 'West Coast', 'Gadget Elite S2', 42, 249.50, 10479.00],
                ['2025-04-12', 'David Park', 'Midwest', 'Widget Pro X1', 65, 129.99, 8449.35],
                ['2025-04-15', 'Rachel Kim', 'West Coast', 'Component Alpha 3', 98, 45.75, 4483.50],
                ['2025-04-17', 'James Wilson', 'Southeast', 'Gadget Elite S2', 36, 249.50, 8982.00],
                ['2025-04-20', 'Sarah Chen', 'Northeast', 'Premium Bundle Z', 21, 599.00, 12579.00],
                ['2025-04-23', 'Marcus Johnson', 'Southeast', 'Widget Pro X1', 57, 129.99, 7409.43],
                ['2025-04-25', 'Emily Rodriguez', 'West Coast', 'Component Alpha 3', 140, 45.75, 6405.00],
                ['2025-04-28', 'David Park', 'Midwest', 'Gadget Elite S2', 30, 249.50, 7485.00],
            ],
        },
        'May': {
            'headers': ['Date', 'Sales Rep', 'Region', 'Product', 'Units Sold', 'Unit Price', 'Total Revenue'],
            'data': [
                ['2025-05-01', 'Rachel Kim', 'West Coast', 'Widget Pro X1', 73, 129.99, 9489.27],
                ['2025-05-03', 'James Wilson', 'Southeast', 'Premium Bundle Z', 20, 599.00, 11980.00],
                ['2025-05-06', 'Sarah Chen', 'Northeast', 'Gadget Elite S2', 46, 249.50, 11477.00],
                ['2025-05-09', 'Marcus Johnson', 'Southeast', 'Component Alpha 3', 115, 45.75, 5261.25],
                ['2025-05-11', 'Emily Rodriguez', 'West Coast', 'Widget Pro X1', 54, 129.99, 7019.46],
                ['2025-05-14', 'David Park', 'Midwest', 'Premium Bundle Z', 16, 599.00, 9584.00],
                ['2025-05-17', 'Rachel Kim', 'West Coast', 'Gadget Elite S2', 40, 249.50, 9980.00],
                ['2025-05-19', 'James Wilson', 'Southeast', 'Widget Pro X1', 68, 129.99, 8839.32],
                ['2025-05-22', 'Sarah Chen', 'Northeast', 'Component Alpha 3', 92, 45.75, 4209.00],
                ['2025-05-25', 'Marcus Johnson', 'Southeast', 'Gadget Elite S2', 37, 249.50, 9231.50],
                ['2025-05-28', 'Emily Rodriguez', 'West Coast', 'Premium Bundle Z', 23, 599.00, 13777.00],
                ['2025-05-30', 'David Park', 'Midwest', 'Widget Pro X1', 60, 129.99, 7799.40],
            ],
        },
    }

    first_sheet = True
    for month_name, month_info in months_data.items():
        if first_sheet:
            ws = wb.active
            ws.title = month_name
            first_sheet = False
        else:
            ws = wb.create_sheet(month_name)

        headers = month_info['headers']
        data = month_info['data']

        # Write headers with styling
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Write data rows
        for r, row_data in enumerate(data, 2):
            for c, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = thin_border
                # Format date column
                if c == 1:
                    cell.number_format = date_fmt
                # Format currency columns
                elif c in (6, 7):
                    cell.number_format = currency_fmt
                # Center-align numeric columns
                elif c in (5,):
                    cell.alignment = Alignment(horizontal="center")

        # Set column widths for readability
        col_widths = [12, 20, 14, 22, 12, 12, 16]
        for i, width in enumerate(col_widths):
            col_letter = chr(65 + i)
            ws.column_dimensions[col_letter].width = width

        # Freeze header row
        ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Open in LibreOffice Calc for GUI-ready state
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
