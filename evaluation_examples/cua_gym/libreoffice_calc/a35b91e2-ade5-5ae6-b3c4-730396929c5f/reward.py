"""
Reward Script: Protect all sheets with password 'secure2024'
Task ID: calc_mcp_017
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): All 5 sheets have protection enabled (0.1 per sheet)
  Component 2 (0.5): All 5 sheets have correct password hash for 'secure2024' (0.1 per sheet)
"""

import os
import openpyxl
from openpyxl.worksheet.protection import SheetProtection

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_017'
EXPECTED_SHEETS = ['Jan', 'Feb', 'Mar', 'Apr', 'May']
PASSWORD = 'secure2024'

# Compute expected password hash once
_ref_prot = SheetProtection(sheet=True, password=PASSWORD)
EXPECTED_HASH = _ref_prot.password


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: expected sheets must exist
    for name in EXPECTED_SHEETS:
        if name not in wb.sheetnames:
            print(f"CRITICAL: Expected sheet '{name}' not found. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0

    # Component 1: Sheet protection enabled (0.5 points — 0.1 per sheet)
    try:
        protected_count = 0
        for name in EXPECTED_SHEETS:
            ws = wb[name]
            if ws.protection.sheet:
                protected_count += 1
                print(f"PASS: Component 1 — Sheet '{name}' is protected")
            else:
                print(f"FAIL: Component 1 — Sheet '{name}' is NOT protected")
        comp1_score = protected_count * 0.1
        if protected_count > 0:
            print(f"PASS: Component 1 — {protected_count}/5 sheets protected ({comp1_score} pts)")
            total_score += comp1_score
        else:
            print(f"FAIL: Component 1 — No sheets protected")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Correct password hash for 'secure2024' (0.5 points — 0.1 per sheet)
    # Only checks sheets that ARE protected (avoids double-counting unprotected sheets)
    try:
        password_match_count = 0
        for name in EXPECTED_SHEETS:
            ws = wb[name]
            if ws.protection.sheet and ws.protection.password == EXPECTED_HASH:
                password_match_count += 1
                print(f"PASS: Component 2 — Sheet '{name}' has correct password hash")
            elif ws.protection.sheet:
                print(f"FAIL: Component 2 — Sheet '{name}' protected but wrong password "
                      f"(expected {EXPECTED_HASH}, got {ws.protection.password})")
            else:
                print(f"FAIL: Component 2 — Sheet '{name}' not protected, cannot verify password")
        comp2_score = password_match_count * 0.1
        if password_match_count > 0:
            print(f"PASS: Component 2 — {password_match_count}/5 sheets with correct password ({comp2_score} pts)")
            total_score += comp2_score
        else:
            print(f"FAIL: Component 2 — No sheets have correct password")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
