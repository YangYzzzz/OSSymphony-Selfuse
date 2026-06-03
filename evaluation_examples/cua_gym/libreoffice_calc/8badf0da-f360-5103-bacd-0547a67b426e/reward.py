"""
Reward Script: Monthly P&L Statement in LibreOffice Calc
Task ID: calc_grs_059
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20): P&L sheet exists and structure (sections present)
  Component 2 (0.20): Revenue line items with 5-column layout
  Component 3 (0.15): COGS, Gross Profit, OpEx, Operating Income, Net Income rows
  Component 4 (0.15): Formulas for variance and totals
  Component 5 (0.10): Bold formatting on section headers and total rows
  Component 6 (0.10): Conditional formatting on variance columns
  Component 7 (0.10): KPI summary box at top
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_059'


def persist_app_state(domain):
    """Best-effort save of any open LibreOffice document."""
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print(f"PERSIST: ctrl+s sent for {domain}")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """Verify P&L statement with progressive scoring. Returns float 0.0-1.0."""
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # ---- Find the P&L sheet (flexible name matching) ----
    pnl_sheet = None
    for name in wb.sheetnames:
        nl = name.lower()
        if 'p&l' in nl or 'profit' in nl or 'loss' in nl or 'pnl' in nl or 'p & l' in nl:
            pnl_sheet = wb[name]
            break
    if pnl_sheet is None:
        # fallback: use the first sheet if it has enough rows
        ws0 = wb.worksheets[0]
        if ws0.max_row >= 20:
            pnl_sheet = ws0

    # Component 1: P&L sheet exists with proper structure (0.20 pts)
    try:
        if pnl_sheet is None:
            print("FAIL: Component 1 — No P&L sheet found")
        else:
            ws = pnl_sheet
            # Must have at least ~30 rows and 5+ columns to hold a full P&L
            if ws.max_row >= 25 and ws.max_column >= 5:
                # Check for key section labels
                labels_found = set()
                required_sections = {'revenue', 'cost', 'gross', 'operating', 'net'}
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            val_lower = cell.value.lower().strip()
                            for sect in required_sections:
                                if sect in val_lower:
                                    labels_found.add(sect)
                if len(labels_found) >= 4:
                    print(f"PASS: Component 1 — P&L sheet found with sections: {labels_found} (0.20 pts)")
                    total_score += 0.20
                else:
                    print(f"FAIL: Component 1 — Only found sections: {labels_found}, need >= 4 of {required_sections}")
            else:
                print(f"FAIL: Component 1 — Sheet too small: {ws.max_row} rows, {ws.max_column} cols")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    if pnl_sheet is None:
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws = pnl_sheet

    # Component 2: Revenue line items with 5-column layout (0.20 pts)
    # Must have: Product Sales, Service Revenue, Subscription Revenue, Other Income
    # Each with Current Month, Prior Month, Variance $, Variance %, YTD columns
    try:
        revenue_items = ['product sales', 'service revenue', 'subscription revenue', 'other income']
        items_found = 0
        items_with_5cols = 0

        for row_idx in range(1, ws.max_row + 1):
            cell_a = ws.cell(row=row_idx, column=1).value
            if cell_a and isinstance(cell_a, str):
                for item in revenue_items:
                    if item in cell_a.lower().strip():
                        items_found += 1
                        # Check that columns B-F have values or formulas
                        cols_filled = 0
                        for col_idx in range(2, 7):
                            v = ws.cell(row=row_idx, column=col_idx).value
                            if v is not None:
                                cols_filled += 1
                        if cols_filled >= 5:
                            items_with_5cols += 1
                        break

        if items_found >= 4 and items_with_5cols >= 3:
            print(f"PASS: Component 2 — {items_found} revenue items found, {items_with_5cols} with 5-col layout (0.20 pts)")
            total_score += 0.20
        elif items_found >= 3 and items_with_5cols >= 2:
            print(f"PARTIAL: Component 2 — {items_found} revenue items, {items_with_5cols} with 5 cols (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 — Only {items_found} revenue items found, {items_with_5cols} with 5-col layout")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: COGS, Gross Profit, OpEx, Operating Income, Below-the-line, Net Income (0.15 pts)
    try:
        key_rows = {
            'cogs': ['cost of goods', 'cogs', 'materials', 'labor', 'overhead'],
            'gross_profit': ['gross profit'],
            'opex': ['operating expense', 'salaries', 'rent', 'utilities', 'marketing', 'r&d', 'research', 'admin', 'depreciation'],
            'operating_income': ['operating income'],
            'below_line': ['interest', 'tax', 'below the line'],
            'net_income': ['net income'],
        }
        sections_found = set()
        for row_idx in range(1, ws.max_row + 1):
            cell_a = ws.cell(row=row_idx, column=1).value
            if cell_a and isinstance(cell_a, str):
                val_lower = cell_a.lower().strip()
                for section, keywords in key_rows.items():
                    for kw in keywords:
                        if kw in val_lower:
                            sections_found.add(section)
                            break

        needed = {'cogs', 'gross_profit', 'opex', 'operating_income', 'net_income'}
        found_needed = sections_found.intersection(needed)
        if len(found_needed) >= 5:
            print(f"PASS: Component 3 — All key P&L sections found: {found_needed} (0.15 pts)")
            total_score += 0.15
        elif len(found_needed) >= 3:
            print(f"PARTIAL: Component 3 — {len(found_needed)}/5 sections found: {found_needed} (0.07 pts)")
            total_score += 0.07
        else:
            print(f"FAIL: Component 3 — Only {len(found_needed)}/5 sections found: {found_needed}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Formulas for variance and totals (0.15 pts)
    # Check that variance columns (D, E) contain formulas and total rows have SUM formulas
    try:
        variance_formulas = 0
        sum_formulas = 0
        pct_formulas = 0

        for row_idx in range(1, ws.max_row + 1):
            # Variance $ (column D)
            d_val = ws.cell(row=row_idx, column=4).value
            if d_val and isinstance(d_val, str) and d_val.startswith('='):
                variance_formulas += 1

            # Variance % (column E)
            e_val = ws.cell(row=row_idx, column=5).value
            if e_val and isinstance(e_val, str) and e_val.startswith('='):
                pct_formulas += 1

            # SUM formulas in col B for totals
            b_val = ws.cell(row=row_idx, column=2).value
            if b_val and isinstance(b_val, str) and '=SUM' in str(b_val).upper():
                sum_formulas += 1

        score_4 = 0.0
        if variance_formulas >= 5:
            score_4 += 0.05
        if pct_formulas >= 5:
            score_4 += 0.05
        if sum_formulas >= 2:
            score_4 += 0.05

        if score_4 > 0:
            print(f"PASS: Component 4 — Variance formulas: {variance_formulas}, Pct formulas: {pct_formulas}, SUM formulas: {sum_formulas} ({score_4} pts)")
            total_score += score_4
        else:
            print(f"FAIL: Component 4 — Insufficient formulas: Var={variance_formulas}, Pct={pct_formulas}, SUM={sum_formulas}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Bold formatting on section headers and total rows (0.10 pts)
    try:
        bold_headers = 0
        bold_total_rows = 0

        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=1)
            if cell.value and isinstance(cell.value, str):
                val_lower = cell.value.lower().strip()
                is_header = any(kw in val_lower for kw in ['revenue', 'cost of goods', 'cogs', 'operating expense', 'below the line'])
                is_total = any(kw in val_lower for kw in ['total', 'gross profit', 'operating income', 'net income'])

                if is_header and cell.font and cell.font.bold:
                    bold_headers += 1
                if is_total and cell.font and cell.font.bold:
                    bold_total_rows += 1

        if bold_headers >= 2 and bold_total_rows >= 2:
            print(f"PASS: Component 5 — {bold_headers} bold headers, {bold_total_rows} bold totals (0.10 pts)")
            total_score += 0.10
        elif bold_headers >= 1 or bold_total_rows >= 1:
            print(f"PARTIAL: Component 5 — {bold_headers} bold headers, {bold_total_rows} bold totals (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — No bold formatting found on headers/totals")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Conditional formatting on variance columns (0.10 pts)
    # Task requires: positive in green, negative in red on variance columns
    try:
        cf_rules = ws.conditional_formatting
        has_green_pos = False
        has_red_neg = False

        for cf in cf_rules:
            cf_range = str(cf)
            for rule in cf.rules:
                rule_type = getattr(rule, 'type', '')
                rule_op = getattr(rule, 'operator', '')
                rule_formula = getattr(rule, 'formula', [])

                # Check for green-on-positive
                if rule_type == 'cellIs' and rule_op == 'greaterThan':
                    has_green_pos = True
                # Check for red-on-negative
                if rule_type == 'cellIs' and rule_op == 'lessThan':
                    has_red_neg = True

        if has_green_pos and has_red_neg:
            print(f"PASS: Component 6 — Conditional formatting found: green>0 and red<0 (0.10 pts)")
            total_score += 0.10
        elif has_green_pos or has_red_neg:
            print(f"PARTIAL: Component 6 — Only partial CF: green>0={has_green_pos}, red<0={has_red_neg} (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 — No conditional formatting on variance columns")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: KPI summary box at top (0.10 pts)
    # Must have Revenue, Gross Margin %, EBITDA, Net Margin % references
    try:
        kpi_found = set()
        kpi_keywords = {
            'revenue_kpi': ['revenue'],
            'gross_margin': ['gross margin', 'gross margin %'],
            'ebitda': ['ebitda'],
            'net_margin': ['net margin', 'net margin %'],
        }
        # Search in the first 6 rows for KPI labels
        for row_idx in range(1, 7):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value and isinstance(cell.value, str):
                    val_lower = cell.value.lower().strip()
                    for kpi, keywords in kpi_keywords.items():
                        for kw in keywords:
                            if kw in val_lower:
                                kpi_found.add(kpi)
                                break

        # Also check that adjacent cells have formulas/values (not just labels)
        has_kpi_values = False
        for row_idx in range(1, 7):
            for col_idx in range(2, ws.max_column + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is not None and isinstance(v, str) and v.startswith('='):
                    has_kpi_values = True
                    break
            if has_kpi_values:
                break

        if len(kpi_found) >= 3 and has_kpi_values:
            print(f"PASS: Component 7 — KPI box found with {kpi_found} and formulas (0.10 pts)")
            total_score += 0.10
        elif len(kpi_found) >= 2:
            print(f"PARTIAL: Component 7 — KPI box partial: {kpi_found} (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 7 — KPI summary box not found (only {kpi_found})")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
