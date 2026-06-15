"""
Reward Script: Paste Special — Values Only from Calculations to Distribution sheet
Task ID: calc_gsi_026
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20): Distribution sheet exists
  Component 2 (0.15): Headers match Calculations sheet
  Component 3 (0.15): Data dimensions correct (99 rows x 4 cols)
  Component 4 (0.25): Column C contains numeric values, not formulas
  Component 5 (0.25): Column D contains numeric values, not formulas
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_026'


def persist_app_state(domain: str):
    """Best-effort save any unsaved GUI state."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify that the Calculations data was pasted as values-only
    into a new Distribution sheet.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook (formula mode — to detect formulas vs values)
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Calculations sheet must exist (this is in both initial and golden)
    if 'Calculations' not in wb.sheetnames:
        print("FAIL: Precondition — 'Calculations' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws_calc = wb['Calculations']

    # Component 1: Distribution sheet exists (0.20 points)
    try:
        if 'Distribution' in wb.sheetnames:
            print("PASS: Component 1 — 'Distribution' sheet exists (0.20 pts)")
            total_score += 0.20
        else:
            print("FAIL: Component 1 — 'Distribution' sheet not found")
            # No Distribution sheet means nothing else to check
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    ws_dist = wb['Distribution']

    # Component 2: Headers match Calculations sheet (0.15 points)
    try:
        calc_headers = [ws_calc.cell(row=1, column=c).value for c in range(1, 5)]
        dist_headers = [ws_dist.cell(row=1, column=c).value for c in range(1, 5)]
        if calc_headers == dist_headers:
            print(f"PASS: Component 2 — Headers match: {dist_headers} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 — Headers mismatch. Calc: {calc_headers}, Dist: {dist_headers}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Data dimensions correct — 99 data rows, 4 columns (0.15 points)
    try:
        dist_max_row = ws_dist.max_row
        dist_max_col = ws_dist.max_column
        calc_max_row = ws_calc.max_row
        # Distribution should have same row count as Calculations (header + 99 data rows = 100)
        row_ok = dist_max_row == calc_max_row
        col_ok = dist_max_col >= 4
        if row_ok and col_ok:
            print(f"PASS: Component 3 — Dimensions correct: {dist_max_row} rows x {dist_max_col} cols (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 — Expected {calc_max_row} rows x 4 cols, found {dist_max_row} rows x {dist_max_col} cols")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Column C in Distribution has numeric values, NOT formulas (0.25 points)
    try:
        formula_count_c = 0
        numeric_count_c = 0
        none_count_c = 0
        for r in range(2, dist_max_row + 1):
            val = ws_dist.cell(row=r, column=3).value
            if isinstance(val, str) and val.startswith('='):
                formula_count_c += 1
            elif isinstance(val, (int, float)):
                numeric_count_c += 1
            elif val is None:
                none_count_c += 1

        if formula_count_c == 0 and numeric_count_c > 0:
            print(f"PASS: Component 4 — Col C has {numeric_count_c} numeric values, 0 formulas (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 4 — Col C: {formula_count_c} formulas, {numeric_count_c} numerics, {none_count_c} None")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Column D in Distribution has numeric values, NOT formulas (0.25 points)
    try:
        formula_count_d = 0
        numeric_count_d = 0
        none_count_d = 0
        for r in range(2, dist_max_row + 1):
            val = ws_dist.cell(row=r, column=4).value
            if isinstance(val, str) and val.startswith('='):
                formula_count_d += 1
            elif isinstance(val, (int, float)):
                numeric_count_d += 1
            elif val is None:
                none_count_d += 1

        if formula_count_d == 0 and numeric_count_d > 0:
            print(f"PASS: Component 5 — Col D has {numeric_count_d} numeric values, 0 formulas (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 5 — Col D: {formula_count_d} formulas, {numeric_count_d} numerics, {none_count_d} None")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
