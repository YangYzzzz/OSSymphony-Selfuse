"""
Initial Setup: Paste Special - values only from formulas
Task ID: calc_gsi_026
Domain: libreoffice_calc

Creates a workbook with:
  - "RawData" sheet: lookup tables and reference rates
  - "Calculations" sheet: columns A-D, rows 1-100 with formulas referencing RawData
The agent must copy A1:D100 from Calculations, then Paste Special > Values Only
onto a new sheet for distribution.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_026'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ========== Sheet 1: RawData ==========
    ws_raw = wb.active
    ws_raw.title = 'RawData'

    # Regional markup rates
    ws_raw['A1'] = 'Region'
    ws_raw['B1'] = 'Markup Rate'
    ws_raw['A1'].font = Font(bold=True)
    ws_raw['B1'].font = Font(bold=True)

    regions = [
        ('Northeast', 0.15),
        ('Southeast', 0.12),
        ('Midwest', 0.10),
        ('Southwest', 0.13),
        ('West Coast', 0.18),
        ('Pacific Northwest', 0.16),
        ('Mountain', 0.11),
        ('Great Lakes', 0.14),
        ('Mid-Atlantic', 0.17),
        ('Deep South', 0.09),
    ]
    for i, (region, rate) in enumerate(regions, 2):
        ws_raw.cell(row=i, column=1, value=region)
        ws_raw.cell(row=i, column=2, value=rate)

    # Product category tax rates
    ws_raw['D1'] = 'Category'
    ws_raw['E1'] = 'Tax Rate'
    ws_raw['D1'].font = Font(bold=True)
    ws_raw['E1'].font = Font(bold=True)

    categories = [
        ('Electronics', 0.085),
        ('Furniture', 0.07),
        ('Clothing', 0.06),
        ('Food & Beverage', 0.04),
        ('Office Supplies', 0.065),
        ('Industrial', 0.075),
        ('Healthcare', 0.05),
        ('Automotive', 0.08),
    ]
    for i, (cat, tax) in enumerate(categories, 2):
        ws_raw.cell(row=i, column=4, value=cat)
        ws_raw.cell(row=i, column=5, value=tax)

    # Shipping cost table
    ws_raw['G1'] = 'Weight Tier'
    ws_raw['H1'] = 'Shipping Cost'
    ws_raw['G1'].font = Font(bold=True)
    ws_raw['H1'].font = Font(bold=True)

    shipping = [
        ('0-5 lbs', 8.50),
        ('5-15 lbs', 14.25),
        ('15-30 lbs', 22.75),
        ('30-50 lbs', 35.00),
        ('50+ lbs', 55.00),
    ]
    for i, (tier, cost) in enumerate(shipping, 2):
        ws_raw.cell(row=i, column=7, value=tier)
        ws_raw.cell(row=i, column=8, value=cost)

    ws_raw.column_dimensions['A'].width = 20
    ws_raw.column_dimensions['B'].width = 14
    ws_raw.column_dimensions['D'].width = 18
    ws_raw.column_dimensions['E'].width = 12
    ws_raw.column_dimensions['G'].width = 14
    ws_raw.column_dimensions['H'].width = 16

    # ========== Sheet 2: Calculations ==========
    ws_calc = wb.create_sheet('Calculations')

    # Headers
    headers = ['Product SKU', 'Base Price ($)', 'Regional Markup ($)', 'Total with Tax ($)']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws_calc.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Product data - 99 rows of data (rows 2-100)
    import random
    random.seed(42)

    product_prefixes = ['ELT', 'FRN', 'CLO', 'FNB', 'OFS', 'IND', 'HLC', 'AUT']
    category_indices = list(range(2, 10))  # rows 2-9 in RawData for category lookup
    region_indices = list(range(2, 12))    # rows 2-11 in RawData for region lookup

    for row in range(2, 101):
        # Column A: Product SKU
        prefix = product_prefixes[(row - 2) % len(product_prefixes)]
        sku = f'{prefix}-{1000 + row - 1:04d}'
        ws_calc.cell(row=row, column=1, value=sku)

        # Column B: Base Price (realistic dollar amounts)
        base_price = round(random.uniform(25.0, 850.0), 2)
        ws_calc.cell(row=row, column=2, value=base_price)
        ws_calc.cell(row=row, column=2).number_format = '$#,##0.00'

        # Column C: Regional Markup = Base Price * VLOOKUP region rate
        # Use a rotating region index to reference RawData
        region_row = region_indices[(row - 2) % len(region_indices)]
        formula_c = f'=B{row}*RawData!B{region_row}'
        ws_calc.cell(row=row, column=3, value=formula_c)
        ws_calc.cell(row=row, column=3).number_format = '$#,##0.00'

        # Column D: Total with Tax = (Base + Markup) * (1 + Tax Rate)
        cat_row = category_indices[(row - 2) % len(category_indices)]
        formula_d = f'=(B{row}+C{row})*(1+RawData!E{cat_row})'
        ws_calc.cell(row=row, column=4, value=formula_d)
        ws_calc.cell(row=row, column=4).number_format = '$#,##0.00'

        # Light alternating row fill
        if row % 2 == 0:
            light_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
            for c in range(1, 5):
                ws_calc.cell(row=row, column=c).fill = light_fill

    ws_calc.column_dimensions['A'].width = 16
    ws_calc.column_dimensions['B'].width = 16
    ws_calc.column_dimensions['C'].width = 20
    ws_calc.column_dimensions['D'].width = 20

    ws_calc.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
