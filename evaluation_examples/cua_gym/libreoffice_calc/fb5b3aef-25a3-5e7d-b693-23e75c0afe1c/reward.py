"""
Reward Script: Build a weekly sales activity report template
Task ID: calc_sales_report_weekly_044
Domain: libreoffice_calc
Scoring:
  Component 1: Row 14 TOTAL row with SUM formulas (0.25)
  Component 2: Row 15 AVERAGE row with AVERAGE formulas (0.25)
  Component 3: Conditional formatting on B3:G13 (red fill, below target) (0.25)
  Component 4: Formatting — Row 2 bold+gray fill, rows 14/15 bold, col G currency (0.25)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_report_weekly_044'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Ensure the sheet exists
    if 'WeeklyActivity' not in wb.sheetnames:
        print("FAIL: Sheet 'WeeklyActivity' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['WeeklyActivity']

    # Component 1: Row 14 TOTAL row — label 'TOTAL' in A14, SUM formulas in B14:G14 (0.25 pts)
    try:
        a14 = ws['A14'].value
        label_ok = a14 is not None and str(a14).strip().upper() == 'TOTAL'

        sum_cols_ok = 0
        sum_cols_total = 6  # B through G
        for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
            cell_val = ws[f'{col_letter}14'].value
            if cell_val is not None:
                val_str = str(cell_val).strip().upper().replace(' ', '')
                if 'SUM' in val_str and '3' in val_str and '13' in val_str:
                    sum_cols_ok += 1

        if label_ok and sum_cols_ok == sum_cols_total:
            print(f"PASS: Component 1 — A14='TOTAL', {sum_cols_ok}/6 SUM formulas in B14:G14 (0.25 pts)")
            total_score += 0.25
        elif label_ok and sum_cols_ok > 0:
            partial = 0.25 * (0.5 + 0.5 * sum_cols_ok / sum_cols_total)
            print(f"PARTIAL: Component 1 — A14 label={'OK' if label_ok else 'MISSING'}, SUM formulas: {sum_cols_ok}/6 ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — A14={repr(a14)}, SUM formulas found: {sum_cols_ok}/6")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Row 15 AVERAGE row — label 'AVERAGE' in A15, AVERAGE formulas in B15:G15 (0.25 pts)
    try:
        a15 = ws['A15'].value
        label_ok = a15 is not None and str(a15).strip().upper() == 'AVERAGE'

        avg_cols_ok = 0
        avg_cols_total = 6  # B through G
        for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
            cell_val = ws[f'{col_letter}15'].value
            if cell_val is not None:
                val_str = str(cell_val).strip().upper().replace(' ', '')
                if 'AVERAGE' in val_str and '3' in val_str and '13' in val_str:
                    avg_cols_ok += 1

        if label_ok and avg_cols_ok == avg_cols_total:
            print(f"PASS: Component 2 — A15='AVERAGE', {avg_cols_ok}/6 AVERAGE formulas in B15:G15 (0.25 pts)")
            total_score += 0.25
        elif label_ok and avg_cols_ok > 0:
            partial = 0.25 * (0.5 + 0.5 * avg_cols_ok / avg_cols_total)
            print(f"PARTIAL: Component 2 — A15 label={'OK' if label_ok else 'MISSING'}, AVERAGE formulas: {avg_cols_ok}/6 ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — A15={repr(a15)}, AVERAGE formulas found: {avg_cols_ok}/6")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Conditional formatting on data range B3:G13 — red fill for below-target values (0.25 pts)
    # Targets: Calls=30, Emails=50, Demos=5, Proposals=3, Deals=1, Revenue=25000
    target_map = {
        'B': 30, 'C': 50, 'D': 5, 'E': 3, 'F': 1, 'G': 25000
    }
    try:
        cf_rules = ws.conditional_formatting
        cf_covered = 0  # how many columns have a valid CF rule

        for col_letter, target_val in target_map.items():
            # Check if there is a CF rule covering this column in range rows 3-13
            col_found = False
            for cf_range in cf_rules:
                cf_str = str(cf_range)
                # Check if this cf_range covers the column (e.g. "B3:B13" or "B3:G13")
                for rule in cf_rules[cf_range]:
                    if rule.type == 'cellIs' and rule.operator == 'lessThan':
                        if rule.formula and str(rule.formula[0]) == str(target_val):
                            # Check that the cf_range includes our column
                            if col_letter in cf_str:
                                col_found = True
                                break
                if col_found:
                    break

            if col_found:
                cf_covered += 1

        # Also check if at least some rules exist and use red fill
        red_fill_found = False
        for cf_range in cf_rules:
            for rule in cf_rules[cf_range]:
                if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                    fill = rule.dxf.fill
                    try:
                        rgb = fill.fgColor.rgb
                        if 'FF0000' in rgb.upper() or rgb.upper() in ('FFFF0000', 'FF0000'):
                            red_fill_found = True
                            break
                    except Exception:
                        pass
            if red_fill_found:
                break

        if cf_covered == 6 and red_fill_found:
            print(f"PASS: Component 3 — Conditional formatting on all 6 columns (B:G) with red fill for below-target values (0.25 pts)")
            total_score += 0.25
        elif cf_covered >= 3:
            partial = 0.25 * (cf_covered / 6)
            print(f"PARTIAL: Component 3 — CF rules for {cf_covered}/6 columns, red_fill={red_fill_found} ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — CF rules for {cf_covered}/6 columns, red_fill={red_fill_found}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Formatting — Row 2 bold+gray fill, rows 14/15 bold, column G $#,##0 format (0.25 pts)
    try:
        format_checks = 0
        format_total = 4  # sub-checks

        # Sub-check 4a: Row 2 (TARGET) bold
        row2_bold = ws['A2'].font.bold or ws['B2'].font.bold
        if row2_bold:
            format_checks += 1
            print(f"  PASS: Row 2 bold formatting present")
        else:
            print(f"  FAIL: Row 2 not bold")

        # Sub-check 4b: Row 2 (TARGET) gray background
        try:
            a2_fill_rgb = ws['A2'].fill.fgColor.rgb
            # Light gray — FFD9D9D9 or similar gray pattern
            is_gray = (a2_fill_rgb not in ('00000000', None) and
                       ws['A2'].fill.fill_type == 'solid' and
                       a2_fill_rgb.upper() != '00000000')
            if is_gray:
                format_checks += 1
                print(f"  PASS: Row 2 gray background present (rgb={a2_fill_rgb})")
            else:
                print(f"  FAIL: Row 2 gray background not found (rgb={a2_fill_rgb}, fill_type={ws['A2'].fill.fill_type})")
        except Exception as fill_e:
            print(f"  ERROR: Row 2 fill check failed: {fill_e}")

        # Sub-check 4c: Rows 14 and 15 bold
        row14_bold = ws['A14'].font.bold or ws['B14'].font.bold
        row15_bold = ws['A15'].font.bold or ws['B15'].font.bold
        if row14_bold and row15_bold:
            format_checks += 1
            print(f"  PASS: Rows 14 and 15 bold formatting present")
        elif row14_bold or row15_bold:
            format_checks += 0.5
            print(f"  PARTIAL: Row 14 bold={row14_bold}, Row 15 bold={row15_bold}")
        else:
            print(f"  FAIL: Rows 14 and 15 not bold")

        # Sub-check 4d: Revenue column G has $#,##0 format (check at least data rows 3-13)
        g_format_ok = 0
        for row in range(2, 16):
            cell = ws.cell(row=row, column=7)
            if cell.value is not None and '$' in (cell.number_format or ''):
                g_format_ok += 1
        if g_format_ok >= 5:
            format_checks += 1
            print(f"  PASS: Revenue column G has currency format ($#,##0) in {g_format_ok} cells")
        else:
            print(f"  FAIL: Revenue column G currency format: only {g_format_ok} cells have $format")

        component_score = 0.25 * (format_checks / format_total)
        if component_score >= 0.24:
            print(f"PASS: Component 4 — All formatting checks passed ({format_checks}/{format_total}) (0.25 pts)")
            total_score += 0.25
        else:
            print(f"PARTIAL: Component 4 — Formatting checks {format_checks}/{format_total} ({component_score:.2f} pts)")
            total_score += component_score

    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
