"""
Initial Setup: Weekly Sales Activity Report Template
Task ID: calc_sales_report_weekly_044
Domain: libreoffice_calc

Creates a WeeklyActivity sheet with:
- Row 1: Headers (Rep, Calls, Emails, Demos, Proposals, Deals Closed, Revenue)
- Row 2: TARGET row with weekly targets
- Rows 3-13: 11 sales reps with realistic activity data
- Rows 14-15: Empty (for TOTAL and AVERAGE to be added by task)

MUST NOT include: conditional formatting, SUM/AVERAGE formulas,
bold formatting on TARGET/TOTAL/AVERAGE rows, gray background, currency format
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_report_weekly_044'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'WeeklyActivity'

    # --- Row 1: Headers ---
    headers = ['Rep', 'Calls', 'Emails', 'Demos', 'Proposals', 'Deals Closed', 'Revenue']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # --- Row 2: TARGET row (no bold, no gray background yet - that's part of the task) ---
    ws.cell(row=2, column=1, value='TARGET')
    targets = [30, 50, 5, 3, 1, 25000]
    for col, val in enumerate(targets, 2):
        ws.cell(row=2, column=col, value=val)

    # --- Rows 3-13: 11 sales reps with realistic activity data ---
    # Some reps hit targets, some miss - makes conditional formatting meaningful
    rep_data = [
        # Rep Name,      Calls, Emails, Demos, Proposals, Deals, Revenue
        ['Alexandra Torres',   35,    58,     6,       4,    2,  42500],
        ['Marcus Williams',    22,    41,     3,       2,    0,  18000],
        ['Priya Kapoor',       38,    62,     7,       5,    3,  67800],
        ['Jordan Lee',         28,    47,     4,       2,    1,  24000],
        ['Samantha Novak',     15,    33,     2,       1,    0,  12500],
        ['David Chen',         41,    55,     8,       6,    2,  58300],
        ['Rachel Kim',         33,    51,     5,       3,    1,  31200],
        ['Brian Okafor',       19,    38,     3,       2,    0,  16700],
        ['Emily Hartman',      44,    66,     9,       5,    3,  72400],
        ['Carlos Mendez',      26,    44,     4,       2,    1,  22800],
        ['Natasha Ivanova',    31,    53,     6,       4,    2,  39600],
    ]

    for r, row_data in enumerate(rep_data, 3):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # --- Rows 14-15: Leave empty (TOTAL and AVERAGE to be filled by task) ---
    # Row 14 and 15 intentionally left blank

    # Set column widths for readability
    ws.column_dimensions['A'].width = 22
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet: WeeklyActivity')
    print(f'  - Row 1: Headers')
    print(f'  - Row 2: TARGET (30, 50, 5, 3, 1, 25000)')
    print(f'  - Rows 3-13: 11 sales reps with activity data')
    print(f'  - Rows 14-15: Empty (for TOTAL and AVERAGE)')


create_initial()
