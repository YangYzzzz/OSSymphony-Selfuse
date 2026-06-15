"""
Reward Script: Delete all conditional formatting rules from A1:H50
Task ID: calc_gfl_029
Domain: libreoffice_calc
Scoring:
  Gate: Data integrity preserved (precondition — no points, early exit if broken)
  Component 1 (0.5): Total CF rule count is zero on Data sheet
  Component 2 (0.5): None of the 5 original CF ranges have rules, no CF in A1:H50
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_029'


def persist_app_state(domain: str):
    """Try to save any unsaved LibreOffice state."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify that all conditional formatting rules have been removed
    while cell data remains intact.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get the Data sheet
    if 'Data' not in wb.sheetnames:
        print("CRITICAL: 'Data' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Data']

    # GATE: Data integrity — cell values must be unchanged (precondition, no points)
    # If data is corrupted, the task was done incorrectly — return 0.0 early.
    try:
        checks = [
            ('A1', 'SKU'),
            ('B1', 'Description'),
            ('C1', 'Stock Level'),
            ('H1', 'Status'),
            ('A2', 'SKU-1000'),
            ('C2', 327),
            ('F2', 20.4),
            ('H2', 'In Stock'),
            ('A50', 'SKU-1048'),
            ('H51', 'Out of Stock'),
        ]
        passed = 0
        for coord, expected in checks:
            val = ws[coord].value
            if isinstance(expected, (int, float)) and val is not None:
                try:
                    if abs(float(val) - expected) < 0.01:
                        passed += 1
                except (ValueError, TypeError):
                    pass
            elif val is not None and str(val).strip() == str(expected).strip():
                passed += 1

        ratio = passed / len(checks)
        if ratio < 0.8:
            print(f"GATE FAIL: Data integrity broken ({passed}/{len(checks)} cells correct) — aborting")
            print("REWARD: 0.0")
            return 0.0
        else:
            print(f"GATE PASS: Data integrity OK ({passed}/{len(checks)} cells correct)")
    except Exception as e:
        print(f"GATE ERROR: Data integrity check failed: {e} — continuing")

    # Component 1: Total conditional formatting rule count is zero (0.5 points)
    # This is the primary task requirement — all CF rules must be removed.
    try:
        cf_list = list(ws.conditional_formatting)
        total_rules = 0
        for cf in cf_list:
            total_rules += len(cf.rules)

        if len(cf_list) == 0 and total_rules == 0:
            print(f"PASS: Component 1 — No conditional formatting rule sets found (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — Found {len(cf_list)} CF rule sets with {total_rules} rules total")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: None of the 5 original CF ranges remain AND no CF in A1:H50 (0.5 points)
    # The initial file had CF rules on: B2:B50, D2:D50, F2:F50, C2:C50, H2:H50
    try:
        original_ranges = {'B2:B50', 'D2:D50', 'F2:F50', 'C2:C50', 'H2:H50'}
        remaining_ranges = set()
        for cf in ws.conditional_formatting:
            range_str = str(cf).replace('<ConditionalFormatting ', '').replace('>', '').strip()
            remaining_ranges.add(range_str)

        overlap = original_ranges.intersection(remaining_ranges)
        if len(overlap) == 0:
            # Also check that no CF rules cover ANY cells in A1:H50 at all
            cf_in_range_count = sum(
                1 for cf in ws.conditional_formatting
                for cell_range in cf.cells.ranges
                if cell_range.min_col <= 8 and cell_range.min_row <= 50
            )

            if cf_in_range_count == 0:
                print(f"PASS: Component 2 — No CF rules in A1:H50, all 5 original ranges cleared (0.5 pts)")
                total_score += 0.5
            else:
                print(f"FAIL: Component 2 — Some CF rules still cover cells within A1:H50")
        else:
            print(f"FAIL: Component 2 — These original CF ranges still have rules: {overlap}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
