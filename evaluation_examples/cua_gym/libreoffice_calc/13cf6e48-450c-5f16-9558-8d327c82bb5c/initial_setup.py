"""
Initial Setup: Delete all conditional formatting rules from range A1:H50
Task ID: calc_gfl_029
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_029'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # --- Headers ---
    headers = ["SKU", "Description", "Stock Level", "Reorder Point",
               "Cost", "Revenue", "Margin", "Status"]
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Column widths ---
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 14

    # --- Product data (50 rows) ---
    product_categories = [
        "Wireless Mouse", "Mechanical Keyboard", "USB-C Hub", "Monitor Stand",
        "Webcam HD", "Noise-Cancelling Headphones", "Laptop Stand", "LED Desk Lamp",
        "Ergonomic Chair Pad", "Cable Management Kit", "External SSD 1TB",
        "Bluetooth Speaker", "Screen Protector", "Stylus Pen", "Power Bank 20000mAh",
        "HDMI Cable 6ft", "Ethernet Adapter", "Portable Charger", "Desk Organizer",
        "USB Flash Drive 64GB", "Wireless Earbuds", "Phone Mount", "Surge Protector",
        "Monitor Arm", "Keyboard Wrist Rest", "Mousepad XL", "Laptop Sleeve 15in",
        "Type-C Cable 3-Pack", "Smart Plug", "Desk Fan", "Ring Light 10in",
        "Docking Station", "Privacy Screen Filter", "Memory Foam Footrest",
        "Wireless Charging Pad", "Document Scanner", "Label Maker", "Air Duster",
        "Cable Clips 50-Pack", "Mini Projector", "Portable Monitor 15.6in",
        "Graphic Tablet", "Noise Machine", "Smart Light Bulb 4-Pack",
        "Webcam Cover 3-Pack", "USB Microphone", "Monitor Riser", "Desk Mat 36x18",
        "Fingerprint Reader", "Laptop Cooling Pad",
    ]

    statuses = ["In Stock", "Low Stock", "Out of Stock", "Backordered", "Discontinued"]
    status_weights = [0.45, 0.25, 0.10, 0.12, 0.08]

    random.seed(42)

    for i in range(50):
        row = i + 2
        sku = f"SKU-{1000 + i:04d}"
        desc = product_categories[i]
        stock = random.randint(0, 500)
        reorder = random.randint(10, 100)
        cost = round(random.uniform(5.0, 250.0), 2)
        revenue = round(cost * random.uniform(1.2, 3.5), 2)
        margin = round((revenue - cost) / revenue * 100, 1)
        status = random.choices(statuses, weights=status_weights, k=1)[0]

        ws.cell(row=row, column=1, value=sku)
        ws.cell(row=row, column=2, value=desc)
        ws.cell(row=row, column=3, value=stock)
        ws.cell(row=row, column=4, value=reorder)
        ws.cell(row=row, column=5, value=cost).number_format = '$#,##0.00'
        ws.cell(row=row, column=6, value=revenue).number_format = '$#,##0.00'
        ws.cell(row=row, column=7, value=margin).number_format = '0.0"%"'
        ws.cell(row=row, column=8, value=status)

    # --- Conditional Formatting Rules ---

    # 1. Color scale on column B (Description length-based) - applied to column B
    #    Using a 3-color scale on column B
    ws.conditional_formatting.add(
        "B2:B50",
        ColorScaleRule(
            start_type="min", start_color="FFF8696B",
            mid_type="percentile", mid_value=50, mid_color="FFFFEB84",
            end_type="max", end_color="FF63BE7B",
        )
    )

    # 2. Color scale on column D (Reorder Point)
    ws.conditional_formatting.add(
        "D2:D50",
        ColorScaleRule(
            start_type="min", start_color="FF63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFFFEB84",
            end_type="max", end_color="FFF8696B",
        )
    )

    # 3. Data bars on column F (Revenue)
    ws.conditional_formatting.add(
        "F2:F50",
        DataBarRule(
            start_type="min", end_type="max",
            color="FF4472C4",
        )
    )

    # 4. Cell value condition on column C (Stock Level) - highlight low stock in red
    red_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
    red_font = Font(color="9C0006")
    ws.conditional_formatting.add(
        "C2:C50",
        CellIsRule(
            operator="lessThan",
            formula=["20"],
            fill=red_fill,
            font=red_font,
        )
    )

    # 5. Cell value condition on column C - highlight high stock in green
    green_fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
    green_font = Font(color="006100")
    ws.conditional_formatting.add(
        "C2:C50",
        CellIsRule(
            operator="greaterThan",
            formula=["200"],
            fill=green_fill,
            font=green_font,
        )
    )

    # 6. Cell value conditions on column H (Status) - highlight "Out of Stock" in red
    ws.conditional_formatting.add(
        "H2:H50",
        CellIsRule(
            operator="equal",
            formula=['"Out of Stock"'],
            fill=red_fill,
            font=red_font,
        )
    )

    # 7. Cell value conditions on column H (Status) - highlight "Discontinued" in orange
    orange_fill = PatternFill(start_color="FFFFCC99", end_color="FFFFCC99", fill_type="solid")
    orange_font = Font(color="BF6000")
    ws.conditional_formatting.add(
        "H2:H50",
        CellIsRule(
            operator="equal",
            formula=['"Discontinued"'],
            fill=orange_fill,
            font=orange_font,
        )
    )

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
