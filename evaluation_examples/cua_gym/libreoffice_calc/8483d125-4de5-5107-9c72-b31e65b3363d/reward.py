"""
Reward Script: Create a pivot table to analyze salary distribution
Task ID: calc_pivot_070
Domain: libreoffice_calc
Scoring:
  Component 1 (0.15): 'Pivot Table' sheet exists (task-introduced)
  Component 2 (0.20): Header row with correct department columns
  Component 3 (0.25): Correct salary range labels as rows
  Component 4 (0.25): Specific ground truth cell values match
  Component 5 (0.15): Grand Total row with correct total (200)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_070'


def persist_app_state(domain):
    """Save any unsaved LibreOffice edits before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            import time
            time.sleep(0.8)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: 'Pivot Table' sheet exists (0.15 points)
    # This is task-introduced — initial file has only 'Payroll'
    try:
        pivot_sheet_name = None
        for sn in wb.sheetnames:
            if 'pivot' in sn.lower():
                pivot_sheet_name = sn
                break
        if pivot_sheet_name is not None:
            print(f"PASS: Component 1 — Pivot table sheet found: '{pivot_sheet_name}' (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — No sheet with 'pivot' in name. Sheets: {wb.sheetnames}")
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws = wb[pivot_sheet_name]

    # Component 2: Header row with correct department columns (0.20 points)
    # Expected columns: Salary Range (or similar), IT, Finance, HR, Marketing, Operations, Grand Total
    try:
        expected_depts = {'it', 'finance', 'hr', 'marketing', 'operations'}
        # Search for the header row (could be row 1, 2, or 3)
        header_row = None
        header_map = {}  # dept_name_lower -> column_index
        salary_col = None
        grand_total_col = None

        for r in range(1, min(ws.max_row + 1, 10)):
            row_vals = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    row_vals.append(str(v).strip().lower())
                else:
                    row_vals.append(None)
            # Check if this row contains at least 3 department names
            found_depts = set()
            for idx, val in enumerate(row_vals):
                if val is not None and val in expected_depts:
                    found_depts.add(val)
            if len(found_depts) >= 3:
                header_row = r
                for idx, val in enumerate(row_vals):
                    col = idx + 1
                    if val is not None:
                        if val in expected_depts:
                            header_map[val] = col
                        elif 'salary' in val or 'range' in val:
                            salary_col = col
                        elif 'grand' in val or 'total' in val:
                            grand_total_col = col
                break

        if header_row is not None and len(header_map) >= 4:
            found_str = ', '.join(sorted(header_map.keys()))
            print(f"PASS: Component 2 — Header row {header_row} has department columns: {found_str} (0.20 pts)")
            total_score += 0.20
        else:
            dept_count = len(header_map) if header_row else 0
            print(f"FAIL: Component 2 — Expected department header columns. Found {dept_count} departments. Header row: {header_row}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Correct salary range labels as rows (0.25 points)
    # Expected ranges: 30000-49999, 50000-69999, 70000-89999, 90000-109999, 110000-129999, 130000-150000
    try:
        expected_ranges = [
            '30000-49999', '50000-69999', '70000-89999',
            '90000-109999', '110000-129999', '130000-150000'
        ]
        # Identify salary_col if not found in header scan
        if salary_col is None and header_row is not None:
            # Assume first column or column before department columns
            salary_col = 1

        found_ranges = []
        range_rows = {}  # range_str -> row_number
        if salary_col is not None and header_row is not None:
            for r in range(header_row + 1, ws.max_row + 1):
                val = ws.cell(row=r, column=salary_col).value
                if val is not None:
                    val_str = str(val).strip().replace(' ', '')
                    # Normalize: remove any formatting differences
                    for er in expected_ranges:
                        if er in val_str or val_str == er:
                            found_ranges.append(er)
                            range_rows[er] = r
                            break

        if len(found_ranges) >= 5:
            print(f"PASS: Component 3 — Found {len(found_ranges)}/6 salary ranges: {found_ranges} (0.25 pts)")
            total_score += 0.25
        elif len(found_ranges) >= 3:
            partial = round(0.25 * len(found_ranges) / 6, 2)
            print(f"PARTIAL: Component 3 — Found {len(found_ranges)}/6 salary ranges ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Found only {len(found_ranges)}/6 expected salary ranges")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Specific ground truth cell values match (0.25 points)
    # From context: 50000-69999/IT=12, 70000-89999/Finance=8
    try:
        checks_passed = 0
        total_checks = 2

        # Check 50000-69999 / IT = 12
        if '50000-69999' in range_rows and 'it' in header_map:
            r = range_rows['50000-69999']
            c = header_map['it']
            val = ws.cell(row=r, column=c).value
            if val is not None and int(val) == 12:
                print(f"PASS: Component 4a — 50000-69999/IT = {val} (expected 12)")
                checks_passed += 1
            else:
                print(f"FAIL: Component 4a — 50000-69999/IT = {val}, expected 12")
        else:
            print(f"FAIL: Component 4a — Could not locate 50000-69999 row or IT column")

        # Check 70000-89999 / Finance = 8
        if '70000-89999' in range_rows and 'finance' in header_map:
            r = range_rows['70000-89999']
            c = header_map['finance']
            val = ws.cell(row=r, column=c).value
            if val is not None and int(val) == 8:
                print(f"PASS: Component 4b — 70000-89999/Finance = {val} (expected 8)")
                checks_passed += 1
            else:
                print(f"FAIL: Component 4b — 70000-89999/Finance = {val}, expected 8")
        else:
            print(f"FAIL: Component 4b — Could not locate 70000-89999 row or Finance column")

        if checks_passed == total_checks:
            print(f"PASS: Component 4 — Both ground truth values match (0.25 pts)")
            total_score += 0.25
        elif checks_passed > 0:
            partial = round(0.25 * checks_passed / total_checks, 2)
            print(f"PARTIAL: Component 4 — {checks_passed}/{total_checks} ground truth checks passed ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — No ground truth values matched")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Grand Total row sums to 200 (0.15 points)
    try:
        gt_row = None
        for r in range(header_row + 1, ws.max_row + 1):
            val = ws.cell(row=r, column=salary_col).value
            if val is not None and 'grand' in str(val).lower() and 'total' in str(val).lower():
                gt_row = r
                break

        if gt_row is not None:
            # Find the grand total value — check the Grand Total column or last column
            if grand_total_col is not None:
                gt_val = ws.cell(row=gt_row, column=grand_total_col).value
            else:
                gt_val = ws.cell(row=gt_row, column=ws.max_column).value
            if gt_val is not None and int(gt_val) == 200:
                print(f"PASS: Component 5 — Grand Total = {gt_val} (expected 200) (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 5 — Grand Total = {gt_val}, expected 200")
        else:
            print("FAIL: Component 5 — Grand Total row not found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
