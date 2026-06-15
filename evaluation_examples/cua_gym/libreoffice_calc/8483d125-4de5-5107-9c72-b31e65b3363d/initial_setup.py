"""
Initial Setup: Create a payroll spreadsheet with 200 employee records
Task ID: calc_pivot_070
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_070'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

# --- Deterministic employee data generation ---
random.seed(42)

DEPARTMENTS = ['IT', 'Finance', 'HR', 'Marketing', 'Operations']
GRADES = ['A', 'B', 'C', 'D']

FIRST_NAMES = [
    'Sarah', 'Marcus', 'Elena', 'James', 'Priya', 'David', 'Mei', 'Carlos',
    'Fatima', 'Alex', 'Yuki', 'Omar', 'Lisa', 'Wei', 'Anna', 'Robert',
    'Nadia', 'Thomas', 'Aisha', 'Daniel', 'Sofia', 'Michael', 'Zara', 'Ryan',
    'Chloe', 'Hassan', 'Julia', 'Kevin', 'Amara', 'Patrick', 'Leila', 'Brian',
    'Nina', 'Samuel', 'Rina', 'George', 'Tanya', 'Victor', 'Megan', 'Ivan',
    'Rachel', 'Hiroshi', 'Clara', 'Derek', 'Sunita', 'Eric', 'Diana', 'Jose',
    'Kira', 'Nathan'
]

LAST_NAMES = [
    'Chen', 'Johnson', 'Petrov', 'Williams', 'Sharma', 'Kim', 'Garcia',
    'Mueller', 'Ahmed', 'O\'Brien', 'Tanaka', 'Silva', 'Anderson', 'Wang',
    'Martinez', 'Singh', 'Brown', 'Nakamura', 'Taylor', 'Ali', 'Wilson',
    'Zhang', 'Davis', 'Patel', 'Thomas', 'Lopez', 'Lee', 'Harris', 'Clark',
    'Nguyen', 'Robinson', 'Walker', 'Young', 'Hall', 'Allen', 'Wright',
    'King', 'Scott', 'Green', 'Baker', 'Adams', 'Nelson', 'Mitchell',
    'Campbell', 'Roberts', 'Turner', 'Phillips', 'Evans', 'Morris', 'Reed'
]

# We need to carefully control the distribution so that:
#   50000-69999 / IT = 12
#   70000-89999 / Finance = 8
#   Grand total = 200
#
# Salary ranges (bins): 30000-49999, 50000-69999, 70000-89999, 90000-109999, 110000-129999, 130000-150000
# Departments: IT, Finance, HR, Marketing, Operations
#
# We'll define a distribution matrix [range][dept] = count, ensuring the constraints.

# Distribution matrix: rows = salary ranges, cols = IT, Finance, HR, Marketing, Operations
# Must sum to 200 total
DISTRIBUTION = {
    (30000, 49999):   {'IT': 5,  'Finance': 6,  'HR': 7,  'Marketing': 8,  'Operations': 6},   # 32
    (50000, 69999):   {'IT': 12, 'Finance': 9,  'HR': 8,  'Marketing': 7,  'Operations': 10},  # 46
    (70000, 89999):   {'IT': 10, 'Finance': 8,  'HR': 9,  'Marketing': 11, 'Operations': 9},   # 47
    (90000, 109999):  {'IT': 8,  'Finance': 7,  'HR': 6,  'Marketing': 5,  'Operations': 7},   # 33
    (110000, 129999): {'IT': 6,  'Finance': 5,  'HR': 4,  'Marketing': 6,  'Operations': 5},   # 26
    (130000, 150000): {'IT': 4,  'Finance': 3,  'HR': 3,  'Marketing': 3,  'Operations': 3},   # 16
}

# Verify total
total = sum(sum(dept_counts.values()) for dept_counts in DISTRIBUTION.values())
assert total == 200, f"Total is {total}, expected 200"

# Generate employees
employees = []
emp_id = 1

for (sal_low, sal_high), dept_counts in DISTRIBUTION.items():
    for dept, count in dept_counts.items():
        for _ in range(count):
            first = random.choice(FIRST_NAMES)
            last = random.choice(LAST_NAMES)
            name = f'{first} {last}'
            salary = random.randint(sal_low, sal_high)
            grade = random.choice(GRADES)
            employees.append((emp_id, name, dept, salary, grade))
            emp_id += 1

# Shuffle to make it look natural (not grouped by range/dept)
random.shuffle(employees)

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Payroll'

    # Headers
    headers = ['EmpID', 'Name', 'Department', 'Salary', 'Grade']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for r, (eid, name, dept, salary, grade) in enumerate(employees, 2):
        ws.cell(row=r, column=1, value=eid)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=dept)
        cell_sal = ws.cell(row=r, column=4, value=salary)
        cell_sal.number_format = '#,##0'
        ws.cell(row=r, column=5, value=grade)

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 8

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
