"""
Initial Setup: Inventory spreadsheet with raw data in Sheet1, empty Sheet2
Task ID: osworld_calc_pivot_multi_styled_005
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_pivot_multi_styled_005'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Inventory ---
    ws1 = wb.active
    ws1.title = 'Sheet1'

    # Headers
    headers = ['Item ID', 'Product Type', 'Supplier', 'Warehouse', 'Stock Quantity', 'Unit Value']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)
        cell = ws1.cell(row=1, column=col)
        cell.font = Font(bold=True)

    # Realistic inventory data
    data = [
        ['ITM-001', 'Electronics', 'TechCorp', 'Warehouse A', 350, 129.99],
        ['ITM-002', 'Clothing',    'FashionHouse', 'Warehouse B', 820, 34.50],
        ['ITM-003', 'Electronics', 'DigiSupply', 'Warehouse C', 210, 299.00],
        ['ITM-004', 'Furniture',   'WoodWorks', 'Warehouse A', 95, 589.00],
        ['ITM-005', 'Clothing',    'TechCorp', 'Warehouse B', 640, 22.75],
        ['ITM-006', 'Food',        'FreshGoods', 'Warehouse D', 1200, 4.99],
        ['ITM-007', 'Electronics', 'TechCorp', 'Warehouse A', 175, 449.00],
        ['ITM-008', 'Furniture',   'HomeStyle', 'Warehouse C', 60, 899.00],
        ['ITM-009', 'Food',        'FreshGoods', 'Warehouse D', 2500, 2.49],
        ['ITM-010', 'Clothing',    'FashionHouse', 'Warehouse A', 430, 54.00],
        ['ITM-011', 'Electronics', 'DigiSupply', 'Warehouse B', 280, 179.99],
        ['ITM-012', 'Furniture',   'WoodWorks', 'Warehouse D', 115, 249.00],
        ['ITM-013', 'Food',        'NutriSource', 'Warehouse B', 980, 8.25],
        ['ITM-014', 'Clothing',    'NutriSource', 'Warehouse C', 510, 47.00],
        ['ITM-015', 'Electronics', 'TechCorp', 'Warehouse D', 320, 219.00],
        ['ITM-016', 'Furniture',   'HomeStyle', 'Warehouse A', 45, 1250.00],
        ['ITM-017', 'Food',        'FreshGoods', 'Warehouse B', 1750, 3.79],
        ['ITM-018', 'Electronics', 'DigiSupply', 'Warehouse C', 190, 349.00],
        ['ITM-019', 'Clothing',    'FashionHouse', 'Warehouse D', 380, 68.00],
        ['ITM-020', 'Furniture',   'WoodWorks', 'Warehouse B', 70, 420.00],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 16
    ws1.column_dimensions['C'].width = 16
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 16
    ws1.column_dimensions['F'].width = 12

    # --- Sheet 2: Empty (no pivot tables, no header) ---
    ws2 = wb.create_sheet('Sheet2')
    # Leave Sheet2 completely empty - agent must create pivot tables here

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched libreoffice --calc with DISPLAY=:0')


create_initial()
