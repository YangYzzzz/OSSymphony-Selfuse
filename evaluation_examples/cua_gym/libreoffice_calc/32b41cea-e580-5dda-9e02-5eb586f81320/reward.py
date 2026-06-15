"""
Reward Script: Create three separate pivot tables in Sheet2 summarizing inventory data
               by warehouse, by product type, and by supplier. Place a merged styled
               header (blue fill, bold white text) spanning the top row above all three pivots.
Task ID: osworld_calc_pivot_multi_styled_005
Domain: libreoffice_calc

Scoring:
  Component 1: Merged styled header in Sheet2 row 1 (blue fill, bold white text, merged A1:F1) — 0.30 pts
  Component 2: Pivot table 1 — Stock Quantity by Warehouse with correct values — 0.25 pts
  Component 3: Pivot table 2 — Stock Quantity by Product Type with correct values — 0.25 pts
  Component 4: Pivot table 3 — Stock Quantity by Supplier with correct values — 0.20 pts
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_pivot_multi_styled_005'


def check_header_style(ws):
    """
    Check if A1 has:
    - A merge range covering A1:F1 (or at least spanning multiple columns)
    - Blue background (fgColor starts with 44 or matches FF4472C4)
    - Bold font
    - White font color (FFFFFFFF or FFFFFF)
    Returns True if all three conditions are met.
    """
    cell = ws['A1']

    # Check merge: B1 should be a MergedCell
    b1 = ws['B1']
    is_merged = isinstance(b1, MergedCell)

    # Check blue fill (accept any blue that contains 4472C4 or similar blue)
    try:
        fg_rgb = cell.fill.fgColor.rgb  # 8-char ARGB
        is_blue = (
            fg_rgb is not None
            and fg_rgb.upper() not in ('00000000', 'FFFFFFFF', '')
            and (
                '4472C4' in fg_rgb.upper()
                or fg_rgb.upper().startswith('FF0000')  # pure blue variants
                or (
                    len(fg_rgb) == 8
                    and int(fg_rgb[2:4], 16) < 100   # R component low
                    and int(fg_rgb[4:6], 16) < 100   # G component low
                    and int(fg_rgb[6:8], 16) > 150   # B component high
                )
            )
        )
    except Exception:
        is_blue = False

    # Check bold font
    is_bold = cell.font.bold is True

    # Check white font color
    try:
        font_rgb = cell.font.color.rgb
        is_white = font_rgb is not None and 'FFFFFF' in font_rgb.upper()
    except Exception:
        is_white = False

    return is_merged, is_blue, is_bold, is_white


def find_pivot_block(ws, header_keyword, start_row=1, end_row=None):
    """
    Search for a row containing header_keyword in column A within the sheet.
    Returns the row number of the header, or None if not found.
    """
    max_r = end_row if end_row else ws.max_row
    for r in range(start_row, max_r + 1):
        cell = ws.cell(row=r, column=1)
        val = cell.value
        if val and isinstance(val, str) and header_keyword.lower() in val.lower():
            return r
    return None


def read_pivot_data(ws, label_row, end_sentinel='Grand Total'):
    """
    Read key/value pairs starting from label_row+1 down to the row containing end_sentinel.
    Returns a dict of {label: value} and the grand total value.
    """
    data = {}
    grand_total = None
    r = label_row + 1
    while r <= ws.max_row:
        key = ws.cell(row=r, column=1).value
        val = ws.cell(row=r, column=2).value
        if key is None:
            break
        if isinstance(key, str) and end_sentinel.lower() in key.lower():
            grand_total = val
            break
        if key is not None:
            data[str(key).strip()] = val
        r += 1
    return data, grand_total


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sheet2 must exist
    if 'Sheet2' not in wb.sheetnames:
        print("FAIL: Sheet2 does not exist in the workbook")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws2 = wb['Sheet2']

    # Precondition: Sheet2 must have meaningful content (not empty)
    if ws2.max_row < 5:
        print("FAIL: Sheet2 appears empty (max_row < 5) — no pivot tables detected")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Component 1: Merged styled header in Sheet2 row 1 (blue fill, bold white text, merged) — 0.30 pts
    try:
        is_merged, is_blue, is_bold, is_white = check_header_style(ws2)
        header_val = ws2['A1'].value

        header_sub_score = 0.0
        if header_val and str(header_val).strip():
            header_sub_score += 0.10
            print(f"PASS: Header has text content: '{header_val}'")
        else:
            print(f"FAIL: Header row A1 has no text (value: {header_val})")

        if is_merged:
            header_sub_score += 0.05
            print("PASS: A1 is part of a merged range (B1 is MergedCell)")
        else:
            print("FAIL: A1 is NOT part of a merged range — B1 is not a MergedCell")

        if is_blue:
            header_sub_score += 0.10
            fg = ws2['A1'].fill.fgColor.rgb
            print(f"PASS: Header has blue background fill: {fg}")
        else:
            try:
                fg = ws2['A1'].fill.fgColor.rgb
            except Exception:
                fg = 'unknown'
            print(f"FAIL: Header does not have blue fill (fgColor={fg})")

        if is_bold:
            header_sub_score += 0.03
            print("PASS: Header font is bold")
        else:
            print("FAIL: Header font is NOT bold")

        if is_white:
            header_sub_score += 0.02
            print("PASS: Header font color is white")
        else:
            try:
                fc = ws2['A1'].font.color.rgb
            except Exception:
                fc = 'unknown'
            print(f"FAIL: Header font color is not white (fontColor={fc})")

        if header_sub_score > 0:
            total_score += header_sub_score
        print(f"Component 1 sub-score: {header_sub_score}/0.30")
    except Exception as e:
        print(f"ERROR: Component 1 (header check): {e}")

    # Component 2: Pivot table 1 — Stock Quantity by Warehouse — 0.25 pts
    try:
        # Find the Warehouse pivot header row (look for "Warehouse" in A column label row)
        wh_header_row = find_pivot_block(ws2, 'Warehouse', start_row=2)
        if wh_header_row is None:
            print("FAIL: Component 2 — Could not find 'Warehouse' pivot table header in Sheet2")
        else:
            # The row with "Warehouse" label is a column header; data follows
            wh_data, wh_total = read_pivot_data(ws2, wh_header_row)

            # Expected warehouse totals from golden file
            expected_warehouses = {
                'Warehouse A': 1095,
                'Warehouse B': 4540,
                'Warehouse C': 970,
                'Warehouse D': 4515,
            }
            expected_grand_total = 11120

            correct_rows = 0
            for wh, expected_qty in expected_warehouses.items():
                actual = wh_data.get(wh)
                if actual is not None:
                    try:
                        if abs(float(actual) - expected_qty) < 0.5:
                            correct_rows += 1
                        else:
                            print(f"FAIL: Warehouse pivot — {wh}: expected {expected_qty}, got {actual}")
                    except (ValueError, TypeError):
                        print(f"FAIL: Warehouse pivot — {wh}: non-numeric value '{actual}'")
                else:
                    print(f"FAIL: Warehouse pivot — {wh} not found in pivot data")

            grand_total_ok = (wh_total is not None and abs(float(wh_total) - expected_grand_total) < 0.5)

            if correct_rows == 4 and grand_total_ok:
                total_score += 0.25
                print(f"PASS: Component 2 — Warehouse pivot has all 4 correct values + grand total {wh_total} (0.25 pts)")
            elif correct_rows >= 2:
                partial = 0.15
                total_score += partial
                print(f"PASS (partial): Component 2 — Warehouse pivot has {correct_rows}/4 correct values ({partial} pts)")
            else:
                print(f"FAIL: Component 2 — Warehouse pivot has only {correct_rows}/4 correct values (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 2 (Warehouse pivot): {e}")

    # Component 3: Pivot table 2 — Stock Quantity by Product Type — 0.25 pts
    try:
        pt_header_row = find_pivot_block(ws2, 'Product Type', start_row=2)
        if pt_header_row is None:
            print("FAIL: Component 3 — Could not find 'Product Type' pivot table header in Sheet2")
        else:
            pt_data, pt_total = read_pivot_data(ws2, pt_header_row)

            # Expected product type totals from golden file
            expected_product_types = {
                'Clothing': 2780,
                'Electronics': 1525,
                'Food': 6430,
                'Furniture': 385,
            }
            expected_grand_total = 11120

            correct_rows = 0
            for pt, expected_qty in expected_product_types.items():
                actual = pt_data.get(pt)
                if actual is not None:
                    try:
                        if abs(float(actual) - expected_qty) < 0.5:
                            correct_rows += 1
                        else:
                            print(f"FAIL: Product Type pivot — {pt}: expected {expected_qty}, got {actual}")
                    except (ValueError, TypeError):
                        print(f"FAIL: Product Type pivot — {pt}: non-numeric value '{actual}'")
                else:
                    print(f"FAIL: Product Type pivot — {pt} not found in pivot data")

            grand_total_ok = (pt_total is not None and abs(float(pt_total) - expected_grand_total) < 0.5)

            if correct_rows == 4 and grand_total_ok:
                total_score += 0.25
                print(f"PASS: Component 3 — Product Type pivot has all 4 correct values + grand total {pt_total} (0.25 pts)")
            elif correct_rows >= 2:
                partial = 0.15
                total_score += partial
                print(f"PASS (partial): Component 3 — Product Type pivot has {correct_rows}/4 correct values ({partial} pts)")
            else:
                print(f"FAIL: Component 3 — Product Type pivot has only {correct_rows}/4 correct values (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 3 (Product Type pivot): {e}")

    # Component 4: Pivot table 3 — Stock Quantity by Supplier — 0.20 pts
    try:
        sup_header_row = find_pivot_block(ws2, 'Supplier', start_row=2)
        if sup_header_row is None:
            print("FAIL: Component 4 — Could not find 'Supplier' pivot table header in Sheet2")
        else:
            sup_data, sup_total = read_pivot_data(ws2, sup_header_row)

            # Expected supplier totals from golden file
            expected_suppliers = {
                'DigiSupply': 680,
                'FashionHouse': 1630,
                'FreshGoods': 5450,
                'HomeStyle': 105,
                'NutriSource': 1490,
                'TechCorp': 1485,
                'WoodWorks': 280,
            }
            expected_grand_total = 11120

            correct_rows = 0
            for sup, expected_qty in expected_suppliers.items():
                actual = sup_data.get(sup)
                if actual is not None:
                    try:
                        if abs(float(actual) - expected_qty) < 0.5:
                            correct_rows += 1
                        else:
                            print(f"FAIL: Supplier pivot — {sup}: expected {expected_qty}, got {actual}")
                    except (ValueError, TypeError):
                        print(f"FAIL: Supplier pivot — {sup}: non-numeric value '{actual}'")
                else:
                    print(f"FAIL: Supplier pivot — {sup} not found in pivot data")

            grand_total_ok = (sup_total is not None and abs(float(sup_total) - expected_grand_total) < 0.5)

            if correct_rows == 7 and grand_total_ok:
                total_score += 0.20
                print(f"PASS: Component 4 — Supplier pivot has all 7 correct values + grand total {sup_total} (0.20 pts)")
            elif correct_rows >= 4:
                partial = 0.12
                total_score += partial
                print(f"PASS (partial): Component 4 — Supplier pivot has {correct_rows}/7 correct values ({partial} pts)")
            else:
                print(f"FAIL: Component 4 — Supplier pivot has only {correct_rows}/7 correct values (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 4 (Supplier pivot): {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
