"""
Initial Setup: Create timesheet with employee time data, no custom headers/footers.
Task ID: calc_mcp_080
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_080'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Timesheet ---
    ws = wb.active
    ws.title = 'Timesheet'

    # Headers
    headers = ['Employee', 'Date', 'Day', 'Start Time', 'End Time',
               'Hours Worked', 'Project', 'Notes']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496',
                              fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows - realistic employee timesheet entries
    data = [
        ['Sarah Chen',       '2025-03-03', 'Monday',    '09:00', '17:30', 8.5,  'Atlas Platform',    'Sprint planning meeting'],
        ['Sarah Chen',       '2025-03-04', 'Tuesday',   '08:45', '17:15', 8.5,  'Atlas Platform',    'API integration work'],
        ['Sarah Chen',       '2025-03-05', 'Wednesday', '09:15', '18:00', 8.75, 'Atlas Platform',    'Code review and testing'],
        ['Marcus Johnson',   '2025-03-03', 'Monday',    '08:30', '17:00', 8.5,  'Beacon Analytics',  'Dashboard redesign'],
        ['Marcus Johnson',   '2025-03-04', 'Tuesday',   '09:00', '17:30', 8.5,  'Beacon Analytics',  'Client feedback session'],
        ['Marcus Johnson',   '2025-03-05', 'Wednesday', '08:45', '16:45', 8.0,  'Horizon CRM',      'Data migration support'],
        ['Elena Rodriguez',  '2025-03-03', 'Monday',    '10:00', '18:30', 8.5,  'Horizon CRM',      'Requirements gathering'],
        ['Elena Rodriguez',  '2025-03-04', 'Tuesday',   '09:30', '18:00', 8.5,  'Horizon CRM',      'Wireframe review'],
        ['Elena Rodriguez',  '2025-03-05', 'Wednesday', '09:00', '17:30', 8.5,  'Atlas Platform',    'Cross-team sync'],
        ['David Park',       '2025-03-03', 'Monday',    '08:00', '16:30', 8.5,  'Pulse Monitoring',  'Incident response drill'],
        ['David Park',       '2025-03-04', 'Tuesday',   '08:15', '17:00', 8.75, 'Pulse Monitoring',  'Alert threshold tuning'],
        ['David Park',       '2025-03-05', 'Wednesday', '09:00', '17:30', 8.5,  'Beacon Analytics',  'Infra provisioning'],
        ['Priya Sharma',     '2025-03-03', 'Monday',    '09:30', '18:00', 8.5,  'Atlas Platform',    'Security audit prep'],
        ['Priya Sharma',     '2025-03-04', 'Tuesday',   '09:00', '17:45', 8.75, 'Atlas Platform',    'Pen-test remediation'],
        ['Priya Sharma',     '2025-03-05', 'Wednesday', '08:30', '17:00', 8.5,  'Horizon CRM',      'Compliance documentation'],
    ]

    data_align = Alignment(horizontal='center', vertical='center')
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c != 8:  # Notes column left-aligned
                cell.alignment = data_align
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')

    # Set column widths for readability
    col_widths = {'A': 20, 'B': 14, 'C': 13, 'D': 13, 'E': 13,
                  'F': 15, 'G': 20, 'H': 28}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # NO custom headers or footers - leave as default

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
