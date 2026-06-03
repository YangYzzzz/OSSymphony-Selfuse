"""
Reward Script: CSV Import with Semicolon Delimiters and UTF-8 Encoding
Task ID: calc_gsi_029
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): xlsx file exists with correct structure (8 cols, 16 rows)
  Component 2 (0.30): Data properly split by semicolons into 8 separate columns
  Component 3 (0.25): UTF-8 special characters preserved correctly
  Component 4 (0.20): Numeric values parsed as numbers (not strings)
"""

import os

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_029'


def verify_task(file_path):
    """
    Verify that the CSV was imported correctly with semicolon delimiters and UTF-8 encoding.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Component 1: File has correct structure — 8 columns, 16 rows (1 header + 15 data) (0.25 pts)
    try:
        max_row = ws.max_row
        max_col = ws.max_column
        # Check we have 8 columns (meaning semicolons were used as delimiters)
        # and at least 16 rows (1 header + 15 data)
        if max_col == 8 and max_row >= 16:
            # Verify headers match expected column names
            expected_headers = [
                'Employee ID', 'Full Name', 'Department', 'City',
                'Quarterly Sales (€)', 'Commission Rate', 'Start Date', 'Notes'
            ]
            actual_headers = [ws.cell(row=1, column=c).value for c in range(1, 9)]
            # Check headers match (allowing for minor variations)
            header_match = all(
                str(ah).strip() == str(eh).strip()
                for ah, eh in zip(actual_headers, expected_headers)
                if ah is not None
            ) and len([h for h in actual_headers if h is not None]) == 8
            if header_match:
                print(f"PASS: Component 1 — 8 columns, {max_row} rows, headers correct (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 1 — headers don't match: {actual_headers}")
        else:
            print(f"FAIL: Component 1 — expected 8 cols and >=16 rows, got {max_col} cols and {max_row} rows")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Data properly split into separate columns (0.30 pts)
    # If semicolons were NOT used as delimiters, all data would be in column A as one string
    # We verify that multiple data rows have values in columns A through H
    try:
        complete_rows = 0
        for r in range(2, min(ws.max_row + 1, 17)):
            vals = [ws.cell(row=r, column=c).value for c in range(1, 9)]
            if all(v is not None for v in vals):
                complete_rows += 1

        # Also check that column A does NOT contain semicolons (would indicate un-split data)
        semicolons_in_colA = sum(
            1 for r in range(2, min(ws.max_row + 1, 17))
            if ws.cell(row=r, column=1).value is not None
            and isinstance(ws.cell(row=r, column=1).value, str)
            and ';' in str(ws.cell(row=r, column=1).value)
        )

        if complete_rows >= 14 and semicolons_in_colA == 0:
            print(f"PASS: Component 2 — {complete_rows} rows with all 8 columns populated, no unsplit semicolons (0.30 pts)")
            total_score += 0.30
        elif semicolons_in_colA > 0:
            print(f"FAIL: Component 2 — semicolons found in column A, data was not split properly")
        else:
            print(f"FAIL: Component 2 — only {complete_rows}/15 rows have all 8 columns populated")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: UTF-8 special characters preserved (0.25 pts)
    # Check specific accented characters that would be garbled without UTF-8
    try:
        utf8_checks_passed = 0
        utf8_checks_total = 5

        # Check a selection of names with special characters
        # Row 2: Sébastien (é)
        val_b2 = str(ws.cell(row=2, column=2).value) if ws.cell(row=2, column=2).value else ''
        if 'é' in val_b2 or 'Sébastien' in val_b2:
            utf8_checks_passed += 1
            print(f"  UTF8 check 1: 'é' in B2 ({val_b2}) — OK")
        else:
            print(f"  UTF8 check 1: 'é' missing in B2 ({repr(val_b2)}) — FAIL")

        # Row 3: Müller (ü)
        val_b3 = str(ws.cell(row=3, column=2).value) if ws.cell(row=3, column=2).value else ''
        if 'ü' in val_b3 or 'Müller' in val_b3:
            utf8_checks_passed += 1
            print(f"  UTF8 check 2: 'ü' in B3 ({val_b3}) — OK")
        else:
            print(f"  UTF8 check 2: 'ü' missing in B3 ({repr(val_b3)}) — FAIL")

        # Row 6: Østergaard (Ø)
        val_b6 = str(ws.cell(row=6, column=2).value) if ws.cell(row=6, column=2).value else ''
        if 'Ø' in val_b6 or 'Østergaard' in val_b6:
            utf8_checks_passed += 1
            print(f"  UTF8 check 3: 'Ø' in B6 ({val_b6}) — OK")
        else:
            print(f"  UTF8 check 3: 'Ø' missing in B6 ({repr(val_b6)}) — FAIL")

        # Row 7: Grażyna (ż)
        val_b7 = str(ws.cell(row=7, column=2).value) if ws.cell(row=7, column=2).value else ''
        if 'ż' in val_b7 or 'Grażyna' in val_b7:
            utf8_checks_passed += 1
            print(f"  UTF8 check 4: 'ż' in B7 ({val_b7}) — OK")
        else:
            print(f"  UTF8 check 4: 'ż' missing in B7 ({repr(val_b7)}) — FAIL")

        # Header row: € symbol in column E header
        val_e1 = str(ws.cell(row=1, column=5).value) if ws.cell(row=1, column=5).value else ''
        if '€' in val_e1:
            utf8_checks_passed += 1
            print(f"  UTF8 check 5: '€' in E1 ({val_e1}) — OK")
        else:
            print(f"  UTF8 check 5: '€' missing in E1 ({repr(val_e1)}) — FAIL")

        if utf8_checks_passed >= 4:
            print(f"PASS: Component 3 — {utf8_checks_passed}/{utf8_checks_total} UTF-8 checks passed (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 3 — only {utf8_checks_passed}/{utf8_checks_total} UTF-8 checks passed")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Numeric values properly parsed as numbers (0.20 pts)
    # Employee IDs should be int/float, Sales/Commission should be float
    try:
        numeric_checks_passed = 0
        numeric_checks_total = 4

        # Check Employee ID in A2 is numeric (1001)
        val_a2 = ws.cell(row=2, column=1).value
        if isinstance(val_a2, (int, float)):
            numeric_checks_passed += 1
            print(f"  Numeric check 1: A2={val_a2} is numeric — OK")
        else:
            print(f"  Numeric check 1: A2={repr(val_a2)} is not numeric — FAIL")

        # Check Quarterly Sales in E2 is numeric (45230.50)
        val_e2 = ws.cell(row=2, column=5).value
        if isinstance(val_e2, (int, float)):
            numeric_checks_passed += 1
            print(f"  Numeric check 2: E2={val_e2} is numeric — OK")
        else:
            print(f"  Numeric check 2: E2={repr(val_e2)} is not numeric — FAIL")

        # Check Commission Rate in F2 is numeric (0.12)
        val_f2 = ws.cell(row=2, column=6).value
        if isinstance(val_f2, (int, float)):
            numeric_checks_passed += 1
            print(f"  Numeric check 3: F2={val_f2} is numeric — OK")
        else:
            print(f"  Numeric check 3: F2={repr(val_f2)} is not numeric — FAIL")

        # Check another sales value E4 is numeric (52100.75)
        val_e4 = ws.cell(row=4, column=5).value
        if isinstance(val_e4, (int, float)):
            numeric_checks_passed += 1
            print(f"  Numeric check 4: E4={val_e4} is numeric — OK")
        else:
            print(f"  Numeric check 4: E4={repr(val_e4)} is not numeric — FAIL")

        if numeric_checks_passed >= 3:
            print(f"PASS: Component 4 — {numeric_checks_passed}/{numeric_checks_total} numeric checks passed (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 4 — only {numeric_checks_passed}/{numeric_checks_total} numeric checks passed")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
