"""
Initial Setup: Competitive analysis matrix in LibreOffice Calc
Task ID: calc_grs_087
Domain: libreoffice_calc

Creates a workbook with:
- Feature Matrix sheet: 25 features x 6 companies, ratings Y/P/N only (no color, no +/-, no tally)
- Competitor Summary sheet: headers only, no content
- Pricing sheet: raw pricing data, no comparison formulas
No conditional formatting, no charts, no score tallies.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_087'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ================================================================
    # Sheet 1: Feature Matrix
    # ================================================================
    ws = wb.active
    ws.title = "Feature Matrix"

    # Column headers
    headers = ["Feature", "Category", "DataFlow Pro (Ours)", "AnalytiX Corp",
               "StreamVault", "InsightHub", "CloudMetrics", "NexGen Data"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Set column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20
    for letter in ["C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[letter].width = 18

    # 25 features organized by category
    features = [
        # Core Features (rows 2-8)
        ("Real-time Data Ingestion", "Core Features", "Y", "Y", "P", "Y", "Y", "P"),
        ("Batch Processing", "Core Features", "Y", "Y", "Y", "P", "Y", "Y"),
        ("Data Transformation Pipelines", "Core Features", "Y", "P", "Y", "Y", "P", "Y"),
        ("Schema Validation", "Core Features", "Y", "Y", "Y", "Y", "N", "P"),
        ("Multi-format Support (CSV/JSON/Parquet)", "Core Features", "Y", "Y", "Y", "P", "Y", "Y"),
        ("Event Stream Processing", "Core Features", "P", "Y", "Y", "N", "Y", "P"),
        ("Data Deduplication", "Core Features", "Y", "P", "N", "Y", "P", "Y"),
        # Advanced Features (rows 9-15)
        ("Machine Learning Integration", "Advanced Features", "P", "Y", "N", "Y", "P", "N"),
        ("Predictive Analytics Dashboard", "Advanced Features", "N", "Y", "N", "Y", "P", "N"),
        ("Natural Language Querying", "Advanced Features", "N", "P", "N", "Y", "N", "N"),
        ("Automated Anomaly Detection", "Advanced Features", "P", "Y", "P", "Y", "N", "P"),
        ("Custom Visualization Builder", "Advanced Features", "Y", "Y", "P", "Y", "Y", "P"),
        ("A/B Testing Framework", "Advanced Features", "N", "P", "N", "P", "Y", "N"),
        ("Real-time Collaboration", "Advanced Features", "Y", "Y", "Y", "P", "Y", "N"),
        # Pricing (rows 16-20)
        ("Free Tier Available", "Pricing", "Y", "N", "Y", "Y", "Y", "N"),
        ("Usage-based Pricing", "Pricing", "Y", "Y", "N", "P", "Y", "Y"),
        ("Enterprise Volume Discounts", "Pricing", "Y", "Y", "Y", "Y", "P", "Y"),
        ("Annual Billing Discount", "Pricing", "P", "Y", "Y", "Y", "Y", "N"),
        ("Custom Enterprise Agreements", "Pricing", "Y", "Y", "P", "Y", "Y", "P"),
        # Support (rows 21-23)
        ("24/7 Phone Support", "Support", "Y", "Y", "N", "P", "Y", "P"),
        ("Dedicated Account Manager", "Support", "Y", "Y", "Y", "Y", "N", "Y"),
        ("Online Knowledge Base", "Support", "Y", "Y", "Y", "Y", "Y", "Y"),
        # Integration (rows 24-26)
        ("REST API Access", "Integration", "Y", "Y", "Y", "Y", "Y", "Y"),
        ("Webhook Support", "Integration", "Y", "Y", "P", "Y", "Y", "N"),
        ("Third-party Connectors (50+)", "Integration", "P", "Y", "Y", "P", "Y", "P"),
    ]

    for r, row_data in enumerate(features, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c >= 3:  # Rating columns
                cell.alignment = Alignment(horizontal="center")

    # Freeze header row
    ws.freeze_panes = "A2"

    # ================================================================
    # Sheet 2: Competitor Summary (headers only, no content)
    # ================================================================
    ws2 = wb.create_sheet("Competitor Summary")
    summary_headers = ["Competitor", "Key Strengths", "Key Weaknesses", "Competitive Threat Level"]
    for col, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 40
    ws2.column_dimensions["C"].width = 40
    ws2.column_dimensions["D"].width = 25

    # Add competitor names only (no analysis content)
    competitors = ["DataFlow Pro (Ours)", "AnalytiX Corp", "StreamVault",
                   "InsightHub", "CloudMetrics", "NexGen Data"]
    for r, name in enumerate(competitors, 2):
        ws2.cell(row=r, column=1, value=name)

    # ================================================================
    # Sheet 3: Pricing Comparison (raw data only)
    # ================================================================
    ws3 = wb.create_sheet("Pricing")
    pricing_headers = ["Company", "Starter ($/mo)", "Professional ($/mo)",
                       "Enterprise ($/mo)", "Free Tier Data Limit"]
    for col, h in enumerate(pricing_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws3.column_dimensions["A"].width = 22
    for letter in ["B", "C", "D", "E"]:
        ws3.column_dimensions[letter].width = 22

    pricing_data = [
        ("DataFlow Pro (Ours)", 49, 199, 799, "5 GB"),
        ("AnalytiX Corp", 79, 299, 1199, "None"),
        ("StreamVault", 39, 149, 599, "10 GB"),
        ("InsightHub", 59, 249, 999, "2 GB"),
        ("CloudMetrics", 29, 179, 699, "15 GB"),
        ("NexGen Data", 89, 349, 1499, "None"),
    ]

    for r, row_data in enumerate(pricing_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            if c in (2, 3, 4) and isinstance(val, (int, float)):
                cell.number_format = '$#,##0'
            if c >= 2:
                cell.alignment = Alignment(horizontal="center")

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
