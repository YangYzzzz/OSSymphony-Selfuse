"""
Reward Script: Competitive Analysis Matrix in LibreOffice Calc
Task ID: calc_grs_087
Domain: libreoffice_calc
Scoring:
  Component 1: Extended rating system (+ and - values present)        — 0.15
  Component 2: Conditional formatting rules for Y/P/N/+/-             — 0.20
  Component 3: Score tally row with COUNTIF formulas                  — 0.15
  Component 4: Competitor Summary filled with strengths/weaknesses    — 0.15
  Component 5: Pricing analysis section added                         — 0.15
  Component 6: Charts sheet with radar chart                          — 0.20
  Total: 1.00
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_087'


def persist_app_state(domain):
    """Try to save any open LibreOffice document before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for", domain)
    except Exception as e:
        print("PERSIST_WARN: save hook failed:", e)


def verify_task(file_path):
    """
    Verify competitive analysis matrix task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Feature Matrix sheet must exist
    if 'Feature Matrix' not in wb.sheetnames:
        print("CRITICAL: 'Feature Matrix' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws_fm = wb['Feature Matrix']

    # ---------------------------------------------------------------
    # Component 1: Extended rating system — + and - values (0.15 pts)
    # The initial file only has Y/P/N. The task requires adding + (superior)
    # and - (inferior) ratings. We check that both + and - appear in the
    # rating cells (C2:H26).
    # ---------------------------------------------------------------
    try:
        plus_count = 0
        minus_count = 0
        for row in ws_fm.iter_rows(min_row=2, max_row=26, min_col=3, max_col=8):
            for cell in row:
                if cell.value == '+':
                    plus_count += 1
                elif cell.value == '-':
                    minus_count += 1

        if plus_count >= 3 and minus_count >= 3:
            print(f"PASS: Component 1 — Extended ratings found: {plus_count} '+' and {minus_count} '-' values (0.15 pts)")
            total_score += 0.15
        elif plus_count >= 1 or minus_count >= 1:
            print(f"PARTIAL: Component 1 — Some extended ratings: {plus_count} '+', {minus_count} '-' (0.075 pts)")
            total_score += 0.075
        else:
            print(f"FAIL: Component 1 — No + or - ratings found in Feature Matrix C2:H26")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ---------------------------------------------------------------
    # Component 2: Conditional formatting rules (0.20 pts)
    # The task requires color-coding: Y=light green, P=light yellow,
    # N=light red, +=dark green, -=-dark red.
    # We verify that conditional formatting rules exist covering
    # the rating area, with rules for at least Y, P, N, +, -.
    # ---------------------------------------------------------------
    try:
        cf_formulas_found = set()
        for cf in ws_fm.conditional_formatting:
            for rule in cf.rules:
                if rule.type == 'cellIs' and rule.operator == 'equal':
                    for f in rule.formula:
                        # Extract the value inside quotes, e.g. '"Y"' -> 'Y'
                        stripped = f.strip().strip('"').strip("'")
                        if stripped in ('Y', 'P', 'N', '+', '-'):
                            cf_formulas_found.add(stripped)

        required = {'Y', 'P', 'N', '+', '-'}
        matched = required.intersection(cf_formulas_found)

        if len(matched) == 5:
            print(f"PASS: Component 2 — All 5 conditional formatting rules found: {matched} (0.20 pts)")
            total_score += 0.20
        elif len(matched) >= 3:
            partial = 0.20 * len(matched) / 5
            print(f"PARTIAL: Component 2 — {len(matched)}/5 CF rules found: {matched} ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {len(matched)}/5 CF rules found: {matched}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---------------------------------------------------------------
    # Component 3: Score tally row with COUNTIF formulas (0.15 pts)
    # The task requires a row counting Y and + values per competitor.
    # In golden, this is row 27 with COUNTIF formulas in C27:H27.
    # We check that a tally row exists (row >= 27) with COUNTIF formulas.
    # ---------------------------------------------------------------
    try:
        tally_found = False
        tally_formula_count = 0
        # Search rows 27-35 for a tally row
        for r in range(27, 36):
            cell_c = ws_fm.cell(row=r, column=3).value
            if cell_c and isinstance(cell_c, str) and 'COUNTIF' in cell_c.upper():
                tally_found = True
                for col in range(3, 9):
                    val = ws_fm.cell(row=r, column=col).value
                    if val and isinstance(val, str) and 'COUNTIF' in val.upper():
                        tally_formula_count += 1
                break

        if tally_found and tally_formula_count >= 5:
            print(f"PASS: Component 3 — Score tally row found with {tally_formula_count} COUNTIF formulas (0.15 pts)")
            total_score += 0.15
        elif tally_found and tally_formula_count >= 2:
            partial = 0.15 * tally_formula_count / 6
            print(f"PARTIAL: Component 3 — Tally row found but only {tally_formula_count}/6 COUNTIF formulas ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No score tally row with COUNTIF formulas found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ---------------------------------------------------------------
    # Component 4: Competitor Summary filled (0.15 pts)
    # Initial file has Competitor Summary headers but empty B/C/D columns.
    # Golden has strengths (B), weaknesses (C), and threat level (D) filled.
    # We check that at least 4 of 6 competitor rows have non-empty B and C.
    # ---------------------------------------------------------------
    try:
        if 'Competitor Summary' not in wb.sheetnames:
            print("FAIL: Component 4 — 'Competitor Summary' sheet not found")
        else:
            ws_cs = wb['Competitor Summary']
            filled_rows = 0
            for r in range(2, 8):
                b_val = ws_cs.cell(row=r, column=2).value
                c_val = ws_cs.cell(row=r, column=3).value
                if b_val and c_val and len(str(b_val)) > 5 and len(str(c_val)) > 5:
                    filled_rows += 1

            if filled_rows >= 5:
                print(f"PASS: Component 4 — {filled_rows}/6 competitor summaries filled (0.15 pts)")
                total_score += 0.15
            elif filled_rows >= 3:
                partial = 0.15 * filled_rows / 6
                print(f"PARTIAL: Component 4 — {filled_rows}/6 competitor summaries filled ({partial:.3f} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 4 — Only {filled_rows}/6 competitor summaries filled")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ---------------------------------------------------------------
    # Component 5: Pricing analysis section (0.15 pts)
    # Initial Pricing sheet has only rows 1-7 (company data).
    # Golden adds analysis rows 9-17 with averages and comparisons.
    # We check for content beyond row 8 in the Pricing sheet,
    # specifically formulas like AVERAGE or references to pricing data.
    # ---------------------------------------------------------------
    try:
        if 'Pricing' not in wb.sheetnames:
            print("FAIL: Component 5 — 'Pricing' sheet not found")
        else:
            ws_pr = wb['Pricing']
            analysis_cells = 0
            has_formula = False
            for r in range(8, 25):
                for c in range(1, 6):
                    val = ws_pr.cell(row=r, column=c).value
                    if val is not None:
                        analysis_cells += 1
                        if isinstance(val, str) and ('AVERAGE' in val.upper() or '=' in val):
                            has_formula = True

            if analysis_cells >= 5 and has_formula:
                print(f"PASS: Component 5 — Pricing analysis section found ({analysis_cells} cells, formulas present) (0.15 pts)")
                total_score += 0.15
            elif analysis_cells >= 3:
                print(f"PARTIAL: Component 5 — Some pricing analysis ({analysis_cells} cells) (0.075 pts)")
                total_score += 0.075
            else:
                print(f"FAIL: Component 5 — No pricing analysis section found beyond row 7 (found {analysis_cells} cells)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # ---------------------------------------------------------------
    # Component 6: Charts sheet with radar chart (0.20 pts)
    # Initial file has no Charts sheet. Golden has a Charts sheet with
    # a RadarChart containing 6 series (one per company).
    # ---------------------------------------------------------------
    try:
        if 'Charts' not in wb.sheetnames:
            print("FAIL: Component 6 — 'Charts' sheet not found")
        else:
            ws_ch = wb['Charts']
            charts = ws_ch._charts
            if len(charts) >= 1:
                chart = charts[0]
                chart_type = type(chart).__name__
                series_count = len(chart.series)

                if 'Radar' in chart_type:
                    if series_count >= 4:
                        print(f"PASS: Component 6 — RadarChart found with {series_count} series (0.20 pts)")
                        total_score += 0.20
                    else:
                        print(f"PARTIAL: Component 6 — RadarChart found but only {series_count} series (0.10 pts)")
                        total_score += 0.10
                else:
                    # Chart exists but not radar type — partial credit
                    print(f"PARTIAL: Component 6 — Chart found ({chart_type}) but not Radar type (0.10 pts)")
                    total_score += 0.10
            else:
                print("FAIL: Component 6 — Charts sheet exists but no charts found")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
