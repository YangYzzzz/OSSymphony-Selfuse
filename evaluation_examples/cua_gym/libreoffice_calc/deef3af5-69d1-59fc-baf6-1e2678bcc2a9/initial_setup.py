"""
Initial Setup: Convert cell comments to cell content
Task ID: calc_tbl_084
Domain: libreoffice_calc

Creates a spreadsheet with employee performance review data.
D2:D11 have comments/notes containing review feedback text.
Column E ("Extracted Feedback") is empty - agent must copy comment text there.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_084'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Performance Reviews"

    # --- Headers ---
    headers = ["Employee", "Department", "Rating", "Score", "Extracted Feedback"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Employee Data ---
    employees = [
        ["Sarah Chen", "Engineering", "Exceeds Expectations", 92],
        ["Marcus Johnson", "Marketing", "Meets Expectations", 78],
        ["Priya Patel", "Finance", "Exceeds Expectations", 95],
        ["David Kim", "Engineering", "Meets Expectations", 81],
        ["Elena Rodriguez", "Human Resources", "Outstanding", 98],
        ["James O'Brien", "Sales", "Needs Improvement", 62],
        ["Aisha Mohammed", "Operations", "Meets Expectations", 75],
        ["Thomas Weber", "Engineering", "Exceeds Expectations", 89],
        ["Yuki Tanaka", "Marketing", "Outstanding", 97],
        ["Robert Okafor", "Finance", "Meets Expectations", 80],
    ]

    # Review feedback comments for D2:D11
    review_comments = [
        "Consistently delivers high-quality code and mentors junior developers effectively.",
        "Solid campaign execution but could improve on data-driven decision making.",
        "Exceptional financial modeling skills; led the Q3 budget optimization project.",
        "Reliable team member who meets deadlines; encourage more proactive communication.",
        "Transformed onboarding process reducing new hire ramp-up time by 40 percent.",
        "Missed two quarterly targets; recommend structured coaching plan for Q2.",
        "Good at process documentation but needs to develop stronger vendor relationships.",
        "Architected the new microservices platform; great technical leadership shown.",
        "Creative campaign ideas drove 25 percent increase in social media engagement.",
        "Accurate reporting and strong attention to detail in monthly close procedures.",
    ]

    data_font = Font(name="Calibri", size=11)
    data_align = Alignment(vertical="center")

    for r, (row_data, comment_text) in enumerate(zip(employees, review_comments), 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

        # Add comment/note to column D (Score)
        score_cell = ws.cell(row=r, column=4)
        score_cell.comment = Comment(comment_text, "HR Manager")

        # Column E is intentionally left empty
        empty_cell = ws.cell(row=r, column=5)
        empty_cell.border = thin_border

    # --- Column Widths ---
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 65

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
