"""
Reward Script: Convert cell comments to actual cell content
Task ID: calc_tbl_084
Domain: libreoffice_calc
Scoring:
  Component 1 (0.6): Comment text correctly extracted into E2:E11
  Component 2 (0.4): Comments removed from D2:D11
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_084'

# Expected comment texts for each row (from the initial comments on D2:D11)
EXPECTED_TEXTS = {
    2: 'Consistently delivers high-quality code and mentors junior developers effectively.',
    3: 'Solid campaign execution but could improve on data-driven decision making.',
    4: 'Exceptional financial modeling skills; led the Q3 budget optimization project.',
    5: 'Reliable team member who meets deadlines; encourage more proactive communication.',
    6: 'Transformed onboarding process reducing new hire ramp-up time by 40 percent.',
    7: 'Missed two quarterly targets; recommend structured coaching plan for Q2.',
    8: 'Good at process documentation but needs to develop stronger vendor relationships.',
    9: 'Architected the new microservices platform; great technical leadership shown.',
    10: 'Creative campaign ideas drove 25 percent increase in social media engagement.',
    11: 'Accurate reporting and strong attention to detail in monthly close procedures.',
}


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import os as _os
    _os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Component 1: Comment text correctly extracted into E2:E11 (0.6 points)
    # Each correct cell gets 0.06 points (10 cells * 0.06 = 0.6)
    try:
        extracted_count = 0
        for row_num in range(2, 12):
            e_val = ws.cell(row=row_num, column=5).value  # Column E
            expected = EXPECTED_TEXTS[row_num]
            if e_val is not None and str(e_val).strip() == expected.strip():
                extracted_count += 1
            else:
                print(f"FAIL: E{row_num} expected '{expected[:50]}...', found: {repr(e_val)}")

        if extracted_count > 0:
            component1_score = round(0.06 * extracted_count, 2)
            total_score += component1_score
            print(f"PASS: Component 1 — {extracted_count}/10 comment texts correctly extracted to column E ({component1_score} pts)")
        else:
            print(f"FAIL: Component 1 — No comment texts found in E2:E11")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Comments removed from D2:D11 (0.4 points)
    # Each removed comment gets 0.04 points (10 cells * 0.04 = 0.4)
    try:
        removed_count = 0
        for row_num in range(2, 12):
            d_cell = ws.cell(row=row_num, column=4)  # Column D
            if d_cell.comment is None:
                removed_count += 1
            else:
                print(f"FAIL: D{row_num} still has comment: '{d_cell.comment.text[:50]}...'")

        if removed_count > 0:
            component2_score = round(0.04 * removed_count, 2)
            total_score += component2_score
            print(f"PASS: Component 2 — {removed_count}/10 comments removed from D2:D11 ({component2_score} pts)")
        else:
            print(f"FAIL: Component 2 — All 10 comments still present on D2:D11")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'

# Attempt to persist unsaved GUI state
persist_app_state("libreoffice_calc")

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
