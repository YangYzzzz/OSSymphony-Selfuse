"""
Initial Setup: Customer Lifetime Value (CLV) calculation sheet - raw data only
Task ID: calc_gpm_093
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_093'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'CLV'

    # --- Title row (just text, no merge/formatting - that's the task) ---
    ws['A1'] = 'Customer Lifetime Value Analysis by Cohort'

    # --- Cohort Retention Table (A3:H9) - raw counts ---
    headers = ['Cohort', 'Month 0', 'Month 1', 'Month 2', 'Month 3', 'Month 4', 'Month 5', 'Month 6']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)

    # Cohort data: initial customers 500-800, decreasing retention
    cohort_data = [
        ['Jan 2025', 750, 638, 555, 480, 413, 355, 302],
        ['Feb 2025', 620, 521, 446, 379, 319, 268, 222],
        ['Mar 2025', 800, 696, 613, 540, 470, 411, 357],
        ['Apr 2025', 550, 457, 385, 323, 268, 219, 178],
        ['May 2025', 680, 585, 510, 442, 382, 330, 281],
        ['Jun 2025', 720, 612, 528, 450, 383, 322, 268],
    ]
    for r, row_data in enumerate(cohort_data, 4):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # --- Retention Rate Table (A11:H17) - raw data, no formulas ---
    for col, h in enumerate(headers, 1):
        ws.cell(row=11, column=col, value=h)

    # Placeholder percentage values (raw numbers, NOT formulas linking to above)
    # These are approximate retention rates the agent should replace with formulas
    retention_data = [
        ['Jan 2025', 1.0, 0.851, 0.740, 0.640, 0.551, 0.473, 0.403],
        ['Feb 2025', 1.0, 0.840, 0.719, 0.611, 0.515, 0.432, 0.358],
        ['Mar 2025', 1.0, 0.870, 0.766, 0.675, 0.588, 0.514, 0.446],
        ['Apr 2025', 1.0, 0.831, 0.700, 0.587, 0.487, 0.398, 0.324],
        ['May 2025', 1.0, 0.860, 0.750, 0.650, 0.562, 0.485, 0.413],
        ['Jun 2025', 1.0, 0.850, 0.733, 0.625, 0.532, 0.447, 0.372],
    ]
    for r, row_data in enumerate(retention_data, 12):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # --- Revenue Per User Table (A19:H25) ---
    for col, h in enumerate(headers, 1):
        ws.cell(row=19, column=col, value=h)

    revenue_data = [
        ['Jan 2025', 25.00, 27.50, 28.00, 30.00, 31.50, 33.00, 35.00],
        ['Feb 2025', 22.00, 24.00, 26.50, 28.00, 29.50, 31.00, 32.50],
        ['Mar 2025', 28.00, 29.00, 30.50, 32.00, 33.00, 34.50, 35.00],
        ['Apr 2025', 20.00, 22.50, 24.00, 26.00, 27.50, 29.00, 30.50],
        ['May 2025', 24.00, 26.00, 28.00, 29.50, 31.00, 32.50, 34.00],
        ['Jun 2025', 26.00, 28.00, 29.50, 31.00, 32.00, 33.50, 35.00],
    ]
    for r, row_data in enumerate(revenue_data, 20):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # --- Row 27: CLV label only, no values ---
    ws.cell(row=27, column=1, value='CLV')

    # --- Row 29: Labels only, no formula ---
    ws.cell(row=29, column=1, value='Average CLV')

    # Set reasonable column widths for readability
    ws.column_dimensions['A'].width = 14
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col_letter].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
