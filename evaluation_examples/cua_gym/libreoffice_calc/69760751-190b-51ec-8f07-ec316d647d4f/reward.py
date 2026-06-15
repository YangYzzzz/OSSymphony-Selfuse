"""
Reward Script: Customer Lifetime Value (CLV) calculation sheet with cohort analysis
Task ID: calc_gpm_093
Domain: libreoffice_calc
Scoring:
  Component 1: Title merged, bold 14pt, white font on dark teal, centered (0.15)
  Component 2: Retention rate formulas reference count table (0.15)
  Component 3: CLV row formulas + Average CLV formula + A29 merged bold 14pt (0.15)
  Component 4: Number formats (#,##0 for counts, 0.0% for rates, $#,##0 for CLV, $#,##0.00 for avg) (0.10)
  Component 5: 3-color scale on C4:H9 (0.10)
  Component 6: Conditional formatting on retention rates (cellIs rules) (0.10)
  Component 7: Data bars on CLV row B27:G27 (0.05)
  Component 8: Thick borders around table sections (0.10)
  Component 9: Two charts (line + bar) with correct titles (0.10)
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_093'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet 'CLV' exists
    if 'CLV' not in wb.sheetnames:
        print("CRITICAL: Sheet 'CLV' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['CLV']

    # Component 1: Title row A1:H1 merged, bold 14pt, white font, dark teal fill, centered (0.15 points)
    try:
        a1 = ws['A1']
        merged_ranges = [str(mr) for mr in ws.merged_cells.ranges]
        is_merged = 'A1:H1' in merged_ranges
        is_bold = a1.font.bold is True
        is_14pt = a1.font.size is not None and abs(a1.font.size - 14.0) < 0.5
        # White font color
        has_white_font = False
        if a1.font.color and a1.font.color.rgb:
            rgb = str(a1.font.color.rgb).upper()
            has_white_font = rgb in ('00FFFFFF', 'FFFFFFFF', 'FFFFFF')
        # Dark teal fill (006666)
        has_teal_fill = False
        if a1.fill.fgColor and a1.fill.fgColor.rgb:
            fill_rgb = str(a1.fill.fgColor.rgb).upper()
            has_teal_fill = '006666' in fill_rgb
        is_centered = a1.alignment.horizontal == 'center'

        sub_checks = [is_merged, is_bold, is_14pt, has_white_font, has_teal_fill, is_centered]
        passed = sum(sub_checks)
        if passed == 6:
            print(f"PASS: Component 1 - Title fully formatted ({passed}/6 sub-checks) (0.15 pts)")
            total_score += 0.15
        elif passed >= 3:
            partial = round(0.15 * passed / 6, 3)
            print(f"PARTIAL: Component 1 - Title {passed}/6 sub-checks (merged={is_merged}, bold={is_bold}, 14pt={is_14pt}, white_font={has_white_font}, teal_fill={has_teal_fill}, centered={is_centered}) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - Title only {passed}/6 sub-checks (merged={is_merged}, bold={is_bold}, 14pt={is_14pt}, white_font={has_white_font}, teal_fill={has_teal_fill}, centered={is_centered})")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Retention rate table has formulas referencing count table (0.15 points)
    # Golden has formulas like =C4/B4, =C5/B5, etc. in C12:H17
    # Initial has raw decimal values (not formulas)
    try:
        formula_count = 0
        expected_formulas = 0
        for row_offset in range(6):  # rows 12-17
            for col in range(3, 9):  # columns C-H
                expected_formulas += 1
                cell = ws.cell(row=12 + row_offset, column=col)
                val = cell.value
                if isinstance(val, str) and val.startswith('=') and '/' in val:
                    formula_count += 1
        if formula_count >= 30:  # all 36 or close
            print(f"PASS: Component 2 - Retention rate formulas present ({formula_count}/{expected_formulas}) (0.15 pts)")
            total_score += 0.15
        elif formula_count >= 15:
            partial = round(0.15 * formula_count / 36, 3)
            print(f"PARTIAL: Component 2 - {formula_count}/{expected_formulas} retention formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 - Only {formula_count}/{expected_formulas} retention formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: CLV row formulas (B27:G27 SUMPRODUCT), D29 AVERAGE, A29:C29 merged, A29 bold 14pt (0.15 points)
    try:
        sub_score = 0.0
        # CLV row formulas
        clv_formula_count = 0
        for col in range(2, 8):  # B-G
            cell = ws.cell(row=27, column=col)
            val = cell.value
            if isinstance(val, str) and val.startswith('=') and 'SUMPRODUCT' in val.upper():
                clv_formula_count += 1
        if clv_formula_count >= 4:
            sub_score += 0.05
            print(f"  CLV formulas: {clv_formula_count}/6 SUMPRODUCT formulas found")

        # D29 AVERAGE formula
        d29 = ws['D29']
        if isinstance(d29.value, str) and '=' in d29.value and 'AVERAGE' in d29.value.upper():
            sub_score += 0.03
            print(f"  D29 AVERAGE formula: {d29.value}")
        else:
            print(f"  D29: expected AVERAGE formula, found {repr(d29.value)}")

        # A29:C29 merged
        merged_ranges = [str(mr) for mr in ws.merged_cells.ranges]
        if 'A29:C29' in merged_ranges:
            sub_score += 0.03
            print(f"  A29:C29 merged: True")
        else:
            print(f"  A29:C29 merged: False (ranges: {merged_ranges})")

        # A29 bold 14pt
        a29 = ws['A29']
        if a29.font.bold is True and a29.font.size is not None and abs(a29.font.size - 14.0) < 0.5:
            sub_score += 0.04
            print(f"  A29 bold 14pt: True")
        else:
            print(f"  A29 bold={a29.font.bold}, size={a29.font.size}")

        if sub_score >= 0.14:
            print(f"PASS: Component 3 - CLV formulas & Average CLV ({sub_score} pts)")
        elif sub_score > 0:
            print(f"PARTIAL: Component 3 - CLV formulas & Average CLV ({sub_score} pts)")
        else:
            print(f"FAIL: Component 3 - No CLV changes detected")
        total_score += sub_score
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Number formats (0.10 points)
    # Count cells (#,##0), retention (0.0%), CLV ($#,##0), avg ($#,##0.00)
    try:
        sub_score = 0.0
        # Check count table number format (#,##0)
        count_fmt_ok = 0
        for coord in ['B4', 'C4', 'H9']:
            cell = ws[coord]
            if cell.number_format and '#,##0' in cell.number_format and '%' not in cell.number_format and '$' not in cell.number_format:
                count_fmt_ok += 1
        if count_fmt_ok >= 2:
            sub_score += 0.025
            print(f"  Count format #,##0: {count_fmt_ok}/3 cells OK")

        # Check retention rate format (0.0%)
        rate_fmt_ok = 0
        for coord in ['C12', 'C13', 'H17']:
            cell = ws[coord]
            if cell.number_format and '%' in cell.number_format:
                rate_fmt_ok += 1
        if rate_fmt_ok >= 2:
            sub_score += 0.025
            print(f"  Rate format 0.0%: {rate_fmt_ok}/3 cells OK")

        # Check CLV format ($#,##0)
        clv_fmt_ok = 0
        for coord in ['B27', 'C27']:
            cell = ws[coord]
            if cell.number_format and '$' in cell.number_format:
                clv_fmt_ok += 1
        if clv_fmt_ok >= 1:
            sub_score += 0.025
            print(f"  CLV format $#,##0: {clv_fmt_ok}/2 cells OK")

        # Check D29 format ($#,##0.00)
        d29 = ws['D29']
        if d29.number_format and '$' in d29.number_format and '00' in d29.number_format:
            sub_score += 0.025
            print(f"  D29 format: {d29.number_format}")

        if sub_score >= 0.09:
            print(f"PASS: Component 4 - Number formats ({sub_score} pts)")
        elif sub_score > 0:
            print(f"PARTIAL: Component 4 - Number formats ({sub_score} pts)")
        else:
            print(f"FAIL: Component 4 - No number formats applied")
        total_score += sub_score
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: 3-color scale on C4:H9 (0.10 points)
    try:
        found_color_scale = False
        for cf in ws.conditional_formatting:
            range_str = str(cf).split()[-1].rstrip('>')  # extract range
            for rule in cf.rules:
                if rule.type == 'colorScale' and rule.colorScale:
                    # Check if range covers the count retention area
                    cf_range = str(cf)
                    if 'C4' in cf_range or 'H9' in cf_range:
                        colors = [getattr(c, 'rgb', '') for c in rule.colorScale.color]
                        if len(colors) >= 3:
                            found_color_scale = True
                            print(f"  3-color scale found on {cf_range} with colors {colors}")
        if found_color_scale:
            print(f"PASS: Component 5 - 3-color scale on retention counts (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 - No 3-color scale found on C4:H9")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    # Component 6: Conditional formatting on retention rates - cellIs rules (0.10 points)
    try:
        found_cell_is = 0
        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            for rule in cf.rules:
                if rule.type == 'cellIs':
                    if 'B12' in cf_range or 'H17' in cf_range or '12' in cf_range:
                        found_cell_is += 1
        if found_cell_is >= 2:
            print(f"PASS: Component 6 - Retention rate conditional formatting ({found_cell_is} cellIs rules) (0.10 pts)")
            total_score += 0.10
        elif found_cell_is >= 1:
            print(f"PARTIAL: Component 6 - Only {found_cell_is} cellIs rule(s) found (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 - No cellIs conditional formatting on retention rates")
    except Exception as e:
        print(f"ERROR: Component 6 - {e}")

    # Component 7: Data bars on CLV row B27:G27 (0.05 points)
    try:
        found_data_bar = False
        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            for rule in cf.rules:
                if rule.type == 'dataBar' and rule.dataBar:
                    if '27' in cf_range:
                        found_data_bar = True
                        print(f"  Data bar found on {cf_range}")
        if found_data_bar:
            print(f"PASS: Component 7 - Data bars on CLV row (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 7 - No data bars found on row 27")
    except Exception as e:
        print(f"ERROR: Component 7 - {e}")

    # Component 8: Thick borders around table sections (0.10 points)
    # Check thick borders on corners of each table section
    try:
        thick_count = 0
        # Table 1 (A3:H9) - check corners
        for coord, sides in [('A3', ['left', 'top']), ('H3', ['right', 'top']),
                              ('A9', ['left', 'bottom']), ('H9', ['right', 'bottom'])]:
            cell = ws[coord]
            for side in sides:
                border_side = getattr(cell.border, side)
                if border_side.style in ('thick', 'medium'):
                    thick_count += 1

        # Table 2 (A11:H17)
        for coord, sides in [('A11', ['left', 'top']), ('H11', ['right', 'top']),
                              ('A17', ['left', 'bottom']), ('H17', ['right', 'bottom'])]:
            cell = ws[coord]
            for side in sides:
                border_side = getattr(cell.border, side)
                if border_side.style in ('thick', 'medium'):
                    thick_count += 1

        # Table 3 (A19:H25)
        for coord, sides in [('A19', ['left', 'top']), ('H19', ['right', 'top']),
                              ('A25', ['left', 'bottom']), ('H25', ['right', 'bottom'])]:
            cell = ws[coord]
            for side in sides:
                border_side = getattr(cell.border, side)
                if border_side.style in ('thick', 'medium'):
                    thick_count += 1

        # 24 total border sides to check (3 tables * 4 corners * 2 sides each)
        if thick_count >= 18:
            print(f"PASS: Component 8 - Thick borders ({thick_count}/24 sides) (0.10 pts)")
            total_score += 0.10
        elif thick_count >= 10:
            partial = round(0.10 * thick_count / 24, 3)
            print(f"PARTIAL: Component 8 - Thick borders ({thick_count}/24 sides) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 8 - Only {thick_count}/24 thick border sides found")
    except Exception as e:
        print(f"ERROR: Component 8 - {e}")

    # Component 9: Two charts - line chart (retention curves) and bar chart (CLV) (0.10 points)
    try:
        charts = ws._charts
        chart_count = len(charts)
        has_line = False
        has_bar = False
        retention_title = False
        clv_title = False

        for c in charts:
            class_name = c.__class__.__name__
            # Extract title text
            title_text = ''
            if c.title and hasattr(c.title, 'tx') and c.title.tx:
                try:
                    rich = c.title.tx.rich
                    if rich and rich.p:
                        for p in rich.p:
                            if p.r:
                                for r in p.r:
                                    title_text += r.t or ''
                except:
                    pass

            if class_name == 'LineChart':
                has_line = True
                if 'retention' in title_text.lower():
                    retention_title = True
            elif class_name == 'BarChart':
                has_bar = True
                if 'lifetime' in title_text.lower() or 'clv' in title_text.lower():
                    clv_title = True

        sub_score = 0.0
        if has_line:
            sub_score += 0.03
        if has_bar:
            sub_score += 0.03
        if retention_title:
            sub_score += 0.02
        if clv_title:
            sub_score += 0.02

        if sub_score >= 0.09:
            print(f"PASS: Component 9 - Charts present (line={has_line}, bar={has_bar}, retention_title={retention_title}, clv_title={clv_title}) ({sub_score} pts)")
        elif sub_score > 0:
            print(f"PARTIAL: Component 9 - Charts (line={has_line}, bar={has_bar}, retention_title={retention_title}, clv_title={clv_title}) ({sub_score} pts)")
        else:
            print(f"FAIL: Component 9 - No charts found (count={chart_count})")
        total_score += sub_score
    except Exception as e:
        print(f"ERROR: Component 9 - {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
