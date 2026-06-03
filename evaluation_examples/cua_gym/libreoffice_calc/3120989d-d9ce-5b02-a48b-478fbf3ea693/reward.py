"""
Reward Script: University Course Workload Planner
Task ID: calc_grs_055
Domain: libreoffice_calc
Scoring:
  Component 1: Total Hours Remaining formulas on course sheets (0.25)
  Component 2: Conditional formatting on due dates (0.20)
  Component 3: Workload chart on Weekly Schedule sheet (0.25)
  Component 4: Grade Projection weighted score formulas (0.30)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_055'

# The 5 course sheet names
COURSE_SHEETS = [
    'CS 601 - Machine Learning',
    'MATH 520 - Statistics',
    'CS 615 - Databases',
    'CS 640 - Networks',
    'PHIL 510 - Ethics in AI',
]

# Grade Projection running grade row locations (golden structure)
# Each course block: header row, column header row, 10 assignment rows, running grade row
# Course 1: rows 1-13, running grade at row 13
# Course 2: rows 15-27, running grade at row 27
# Course 3: rows 29-41, running grade at row 41
# Course 4: rows 43-55, running grade at row 55
# Course 5: rows 57-69, running grade at row 69
GRADE_PROJECTION_BLOCKS = [
    {'first_data_row': 3, 'last_data_row': 12, 'running_grade_row': 13},
    {'first_data_row': 17, 'last_data_row': 26, 'running_grade_row': 27},
    {'first_data_row': 31, 'last_data_row': 40, 'running_grade_row': 41},
    {'first_data_row': 45, 'last_data_row': 54, 'running_grade_row': 55},
    {'first_data_row': 59, 'last_data_row': 68, 'running_grade_row': 69},
]


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Total Hours Remaining formulas on each course sheet (0.25 points)
    # Each course sheet should have a SUMPRODUCT formula in the D column (row 12+)
    # that calculates remaining hours for incomplete assignments.
    # This is NOT present in the initial file.
    try:
        comp1_score = 0.0
        sheets_with_formula = 0
        for sn in COURSE_SHEETS:
            if sn not in wb.sheetnames:
                print(f"FAIL: Component 1 — sheet '{sn}' not found")
                continue
            ws = wb[sn]
            # Search rows 12-20 for a "Total Hours Remaining" formula
            found_formula = False
            for r in range(12, min(ws.max_row + 1, 25)):
                cell_a = ws.cell(row=r, column=1).value
                cell_d = ws.cell(row=r, column=4).value
                if cell_a and 'total' in str(cell_a).lower() and 'remaining' in str(cell_a).lower():
                    if cell_d and isinstance(cell_d, str) and '=' in cell_d:
                        # Has a formula for hours remaining
                        found_formula = True
                        sheets_with_formula += 1
                        break
                    elif cell_d and isinstance(cell_d, (int, float)):
                        # Might be a cached computed value — still counts
                        found_formula = True
                        sheets_with_formula += 1
                        break
            if not found_formula:
                print(f"FAIL: Component 1 — no Total Hours Remaining formula found in sheet '{sn}'")

        if sheets_with_formula == 5:
            comp1_score = 0.25
            print(f"PASS: Component 1 — Total Hours Remaining formula found in all 5 course sheets (0.25 pts)")
        elif sheets_with_formula >= 3:
            comp1_score = 0.15
            print(f"PARTIAL: Component 1 — formula found in {sheets_with_formula}/5 sheets (0.15 pts)")
        elif sheets_with_formula >= 1:
            comp1_score = 0.05
            print(f"PARTIAL: Component 1 — formula found in {sheets_with_formula}/5 sheets (0.05 pts)")
        else:
            print(f"FAIL: Component 1 — no Total Hours Remaining formulas found in any course sheet")
        total_score += comp1_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Conditional formatting on due dates within 7 days in red (0.20 points)
    # Each course sheet should have conditional formatting on B column (due dates)
    # with red highlighting for dates within 7 days.
    # Initial file has NO conditional formatting on any course sheet.
    try:
        comp2_score = 0.0
        sheets_with_cf = 0
        for sn in COURSE_SHEETS:
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            cf_list = list(ws.conditional_formatting)
            has_date_cf = False
            for cf in cf_list:
                for rule in cf.rules:
                    # Check if the rule involves date-based or TODAY()-based formula
                    if rule.formula:
                        formula_str = str(rule.formula).upper()
                        if 'TODAY' in formula_str or 'DATE' in formula_str:
                            has_date_cf = True
                            break
                    # Also check for CellIsRule type conditions on date columns
                    if rule.type == 'expression' and rule.formula:
                        formula_str = str(rule.formula).upper()
                        if 'B' in formula_str:
                            has_date_cf = True
                            break
                if has_date_cf:
                    break
            if has_date_cf:
                sheets_with_cf += 1

        if sheets_with_cf == 5:
            comp2_score = 0.20
            print(f"PASS: Component 2 — Conditional formatting on due dates found in all 5 course sheets (0.20 pts)")
        elif sheets_with_cf >= 3:
            comp2_score = 0.12
            print(f"PARTIAL: Component 2 — Conditional formatting found in {sheets_with_cf}/5 sheets (0.12 pts)")
        elif sheets_with_cf >= 1:
            comp2_score = 0.05
            print(f"PARTIAL: Component 2 — Conditional formatting found in {sheets_with_cf}/5 sheets (0.05 pts)")
        else:
            print(f"FAIL: Component 2 — no conditional formatting found on due date columns")
        total_score += comp2_score
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Workload chart on Weekly Schedule sheet (0.25 points)
    # The Weekly Schedule sheet should have at least 1 chart showing estimated hours per week.
    # Initial file has 0 charts on the Weekly Schedule sheet.
    try:
        comp3_score = 0.0
        if 'Weekly Schedule' not in wb.sheetnames:
            print("FAIL: Component 3 — 'Weekly Schedule' sheet not found")
        else:
            ws_sched = wb['Weekly Schedule']
            charts = ws_sched._charts
            if len(charts) >= 1:
                chart = charts[0]
                # Verify it has multiple data series (one per course)
                series_count = len(chart.series)
                if series_count >= 3:
                    comp3_score = 0.25
                    print(f"PASS: Component 3 — Chart found on Weekly Schedule with {series_count} series (0.25 pts)")
                else:
                    comp3_score = 0.15
                    print(f"PARTIAL: Component 3 — Chart found but only {series_count} series (expected >=3) (0.15 pts)")
            else:
                print(f"FAIL: Component 3 — no charts found on Weekly Schedule sheet")
        total_score += comp3_score
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Grade Projection weighted score formulas and running grade (0.30 points)
    # In the golden file, column D has IF formulas for weighted scores, and there are
    # running course grade formulas. In the initial file, column D has NO formulas.
    try:
        comp4_score = 0.0
        if 'Grade Projection' not in wb.sheetnames:
            print("FAIL: Component 4 — 'Grade Projection' sheet not found")
        else:
            ws_gp = wb['Grade Projection']
            blocks_with_d_formulas = 0
            blocks_with_running_grade = 0

            for block in GRADE_PROJECTION_BLOCKS:
                fdr = block['first_data_row']
                ldr = block['last_data_row']
                rgr = block['running_grade_row']

                # Check if D column has formulas in data rows
                d_formula_count = 0
                for r in range(fdr, ldr + 1):
                    val = ws_gp.cell(row=r, column=4).value
                    if val and isinstance(val, str) and '=' in val:
                        d_formula_count += 1
                if d_formula_count >= 5:  # at least half of the assignment rows
                    blocks_with_d_formulas += 1

                # Check running grade formula
                rg_val = ws_gp.cell(row=rgr, column=3).value
                if rg_val and isinstance(rg_val, str) and '=' in str(rg_val):
                    blocks_with_running_grade += 1

            # Score: D formulas worth 0.15, running grade formulas worth 0.15
            d_sub = 0.0
            if blocks_with_d_formulas == 5:
                d_sub = 0.15
                print(f"PASS: Component 4a — Weighted score formulas in D column for all 5 course blocks (0.15 pts)")
            elif blocks_with_d_formulas >= 3:
                d_sub = 0.09
                print(f"PARTIAL: Component 4a — Weighted score formulas found in {blocks_with_d_formulas}/5 blocks (0.09 pts)")
            elif blocks_with_d_formulas >= 1:
                d_sub = 0.03
                print(f"PARTIAL: Component 4a — Weighted score formulas found in {blocks_with_d_formulas}/5 blocks (0.03 pts)")
            else:
                print(f"FAIL: Component 4a — no weighted score formulas found in Grade Projection D column")

            rg_sub = 0.0
            if blocks_with_running_grade == 5:
                rg_sub = 0.15
                print(f"PASS: Component 4b — Running course grade formulas for all 5 courses (0.15 pts)")
            elif blocks_with_running_grade >= 3:
                rg_sub = 0.09
                print(f"PARTIAL: Component 4b — Running grade formulas found in {blocks_with_running_grade}/5 blocks (0.09 pts)")
            elif blocks_with_running_grade >= 1:
                rg_sub = 0.03
                print(f"PARTIAL: Component 4b — Running grade formulas found in {blocks_with_running_grade}/5 blocks (0.03 pts)")
            else:
                print(f"FAIL: Component 4b — no running course grade formulas found")

            comp4_score = d_sub + rg_sub
        total_score += comp4_score
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
