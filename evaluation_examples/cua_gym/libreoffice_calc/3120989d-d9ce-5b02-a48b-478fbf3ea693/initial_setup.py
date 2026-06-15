"""
Initial Setup: University Course Workload Planner
Task ID: calc_grs_055
Domain: libreoffice_calc

Creates a semester workload planner with 5 course sheets, a weekly schedule grid,
and a grade projection section. Does NOT include: conditional formatting on due dates,
total workload chart, or total hours remaining formulas.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_055'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


# --- Style definitions ---
header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

# Course definitions with realistic graduate-level content
COURSES = {
    "CS 601 - Machine Learning": [
        ("Homework 1: Linear Regression", "2026-01-27", 8, 12, "Y"),
        ("Quiz 1: Probability Review", "2026-02-03", 5, 4, "Y"),
        ("Homework 2: Logistic Regression & SVMs", "2026-02-17", 8, 14, "Y"),
        ("Midterm Project Proposal", "2026-02-24", 5, 8, "Y"),
        ("Homework 3: Neural Networks", "2026-03-10", 8, 16, "N"),
        ("Quiz 2: Optimization Methods", "2026-03-17", 5, 4, "N"),
        ("Homework 4: CNNs and Transfer Learning", "2026-03-31", 8, 18, "N"),
        ("Midterm Exam", "2026-04-07", 15, 20, "N"),
        ("Final Project: Model Implementation", "2026-05-05", 25, 40, "N"),
        ("Final Exam", "2026-05-12", 13, 22, "N"),
    ],
    "MATH 520 - Statistics": [
        ("Problem Set 1: Descriptive Statistics", "2026-01-29", 7, 8, "Y"),
        ("Problem Set 2: Probability Distributions", "2026-02-12", 7, 10, "Y"),
        ("Quiz 1: Hypothesis Testing Basics", "2026-02-19", 5, 3, "Y"),
        ("Problem Set 3: Regression Analysis", "2026-03-05", 7, 12, "Y"),
        ("Midterm Exam", "2026-03-19", 20, 18, "N"),
        ("Problem Set 4: Bayesian Inference", "2026-04-02", 7, 14, "N"),
        ("Problem Set 5: Time Series Analysis", "2026-04-16", 7, 12, "N"),
        ("Quiz 2: ANOVA and Experimental Design", "2026-04-23", 5, 4, "N"),
        ("Research Paper Analysis", "2026-05-01", 15, 20, "N"),
        ("Final Exam", "2026-05-14", 20, 22, "N"),
    ],
    "CS 615 - Databases": [
        ("Lab 1: ER Diagram Design", "2026-01-30", 5, 6, "Y"),
        ("Homework 1: Relational Algebra", "2026-02-06", 8, 10, "Y"),
        ("Lab 2: SQL Queries", "2026-02-20", 5, 8, "Y"),
        ("Homework 2: Normalization", "2026-03-06", 8, 12, "Y"),
        ("Midterm Exam", "2026-03-13", 20, 16, "N"),
        ("Lab 3: Transaction Management", "2026-03-27", 5, 8, "N"),
        ("Homework 3: Query Optimization", "2026-04-10", 8, 14, "N"),
        ("Group Project: Database Application", "2026-04-24", 20, 35, "N"),
        ("Lab 4: NoSQL Implementation", "2026-05-01", 5, 8, "N"),
        ("Final Exam", "2026-05-15", 16, 20, "N"),
    ],
    "CS 640 - Networks": [
        ("Lab 1: Wireshark Packet Analysis", "2026-01-28", 5, 5, "Y"),
        ("Homework 1: Application Layer Protocols", "2026-02-11", 8, 10, "Y"),
        ("Lab 2: Socket Programming", "2026-02-25", 5, 10, "Y"),
        ("Homework 2: Transport Layer", "2026-03-04", 8, 12, "N"),
        ("Midterm Exam", "2026-03-18", 20, 18, "N"),
        ("Lab 3: Routing Simulation", "2026-04-01", 5, 8, "N"),
        ("Homework 3: Network Security", "2026-04-15", 8, 14, "N"),
        ("Lab 4: SDN with OpenFlow", "2026-04-29", 5, 10, "N"),
        ("Research Presentation", "2026-05-06", 16, 15, "N"),
        ("Final Exam", "2026-05-13", 20, 20, "N"),
    ],
    "PHIL 510 - Ethics in AI": [
        ("Reading Response 1: Turing and Consciousness", "2026-02-04", 5, 4, "Y"),
        ("Reading Response 2: Algorithmic Bias", "2026-02-18", 5, 4, "Y"),
        ("Essay 1: Autonomous Weapons Ethics", "2026-03-04", 15, 18, "Y"),
        ("Reading Response 3: Privacy and Surveillance", "2026-03-18", 5, 4, "N"),
        ("Group Debate: AI Regulation", "2026-03-25", 10, 8, "N"),
        ("Reading Response 4: Labor Displacement", "2026-04-08", 5, 4, "N"),
        ("Essay 2: AI Alignment Problem", "2026-04-22", 15, 20, "N"),
        ("Reading Response 5: Digital Personhood", "2026-05-01", 5, 4, "N"),
        ("Final Paper: Comprehensive AI Ethics Framework", "2026-05-08", 25, 30, "N"),
        ("Class Participation", "2026-05-15", 10, 0, "N"),
    ],
}

# Semester weeks for weekly schedule
SEMESTER_START = datetime(2026, 1, 26)
NUM_WEEKS = 16


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def create_initial():
    wb = openpyxl.Workbook()

    # ============================================================
    # Create individual course sheets
    # ============================================================
    first = True
    for course_name, assignments in COURSES.items():
        if first:
            ws = wb.active
            ws.title = course_name
            first = False
        else:
            ws = wb.create_sheet(course_name)

        # Headers
        headers = ["Assignment Name", "Due Date", "Weight %", "Estimated Hours", "Completed (Y/N)"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        style_header_row(ws, 1, len(headers))

        # Column widths
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 16

        # Data rows
        for r, (name, due, weight, hours, completed) in enumerate(assignments, 2):
            ws.cell(row=r, column=1, value=name)
            ws.cell(row=r, column=2, value=due)
            ws.cell(row=r, column=2).number_format = 'yyyy-mm-dd'
            ws.cell(row=r, column=3, value=weight)
            ws.cell(row=r, column=3).number_format = '0"%"'
            ws.cell(row=r, column=4, value=hours)
            ws.cell(row=r, column=5, value=completed)
            # Style data cells
            for col in range(1, 6):
                cell = ws.cell(row=r, column=col)
                cell.border = thin_border
                if col >= 2:
                    cell.alignment = Alignment(horizontal="center")

        # Freeze header
        ws.freeze_panes = "A2"

    # ============================================================
    # Weekly Schedule sheet
    # ============================================================
    ws_sched = wb.create_sheet("Weekly Schedule")
    # Headers: Week | Dates | Course1 Hours | Course2 Hours | ... | Total
    course_names_short = ["CS601-ML", "MATH520", "CS615-DB", "CS640-Net", "PHIL510"]
    sched_headers = ["Week", "Date Range"] + course_names_short + ["Total Hours"]
    for col, h in enumerate(sched_headers, 1):
        ws_sched.cell(row=1, column=col, value=h)
    style_header_row(ws_sched, 1, len(sched_headers))

    ws_sched.column_dimensions["A"].width = 8
    ws_sched.column_dimensions["B"].width = 26
    for i in range(3, 3 + len(course_names_short)):
        ws_sched.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 14
    ws_sched.column_dimensions[openpyxl.utils.get_column_letter(len(sched_headers))].width = 14

    # Distribute estimated hours across weeks based on due dates
    # Simple model: hours spread over 2 weeks before due date
    course_list = list(COURSES.keys())
    weekly_hours = [[0.0] * len(course_list) for _ in range(NUM_WEEKS)]

    for ci, (course_name, assignments) in enumerate(COURSES.items()):
        for name, due_str, weight, hours, completed in assignments:
            due = datetime.strptime(due_str, "%Y-%m-%d")
            due_week = max(0, min(NUM_WEEKS - 1, (due - SEMESTER_START).days // 7))
            # Spread hours: 60% in due week, 40% in previous week
            weekly_hours[due_week][ci] += hours * 0.6
            if due_week > 0:
                weekly_hours[due_week - 1][ci] += hours * 0.4

    for w in range(NUM_WEEKS):
        row = w + 2
        week_start = SEMESTER_START + timedelta(weeks=w)
        week_end = week_start + timedelta(days=6)
        ws_sched.cell(row=row, column=1, value=w + 1)
        ws_sched.cell(row=row, column=2, value=f"{week_start.strftime('%b %d')} - {week_end.strftime('%b %d')}")
        total = 0
        for ci in range(len(course_list)):
            val = round(weekly_hours[w][ci], 1)
            ws_sched.cell(row=row, column=3 + ci, value=val)
            ws_sched.cell(row=row, column=3 + ci).alignment = Alignment(horizontal="center")
            ws_sched.cell(row=row, column=3 + ci).border = thin_border
            total += val
        ws_sched.cell(row=row, column=len(sched_headers), value=round(total, 1))
        ws_sched.cell(row=row, column=len(sched_headers)).alignment = Alignment(horizontal="center")
        ws_sched.cell(row=row, column=len(sched_headers)).border = thin_border
        # Style row
        ws_sched.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws_sched.cell(row=row, column=1).border = thin_border
        ws_sched.cell(row=row, column=2).border = thin_border

    ws_sched.freeze_panes = "C2"

    # ============================================================
    # Grade Projection sheet (structure only, no formulas yet)
    # ============================================================
    ws_grade = wb.create_sheet("Grade Projection")

    row = 1
    for ci, (course_name, assignments) in enumerate(COURSES.items()):
        # Course header
        ws_grade.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws_grade.cell(row=row, column=1, value=course_name)
        ws_grade.cell(row=row, column=1).font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
        ws_grade.cell(row=row, column=1).fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
        ws_grade.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        row += 1

        # Sub-headers
        grade_headers = ["Assignment", "Weight %", "Expected Score", "Weighted Score", "Status"]
        for col, h in enumerate(grade_headers, 1):
            cell = ws_grade.cell(row=row, column=col, value=h)
            cell.font = Font(name="Arial", size=10, bold=True)
            cell.fill = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        row += 1

        # Assignment rows (no formulas - those go in golden)
        for name, due, weight, hours, completed in assignments:
            ws_grade.cell(row=row, column=1, value=name)
            ws_grade.cell(row=row, column=2, value=weight)
            ws_grade.cell(row=row, column=2).number_format = '0"%"'
            ws_grade.cell(row=row, column=2).alignment = Alignment(horizontal="center")
            # Expected Score column left blank for user input
            ws_grade.cell(row=row, column=3, value=None)
            # Weighted Score column left blank (formula goes in golden)
            ws_grade.cell(row=row, column=4, value=None)
            ws_grade.cell(row=row, column=5, value="Completed" if completed == "Y" else "Pending")
            for col in range(1, 6):
                ws_grade.cell(row=row, column=col).border = thin_border
            row += 1

        # Running Grade row (placeholder, no formula)
        ws_grade.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws_grade.cell(row=row, column=1, value="Running Course Grade:")
        ws_grade.cell(row=row, column=1).font = Font(bold=True)
        ws_grade.cell(row=row, column=3, value=None)
        for col in range(1, 6):
            ws_grade.cell(row=row, column=col).border = thin_border
        row += 2  # blank row between courses

    ws_grade.column_dimensions["A"].width = 40
    ws_grade.column_dimensions["B"].width = 12
    ws_grade.column_dimensions["C"].width = 16
    ws_grade.column_dimensions["D"].width = 16
    ws_grade.column_dimensions["E"].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
