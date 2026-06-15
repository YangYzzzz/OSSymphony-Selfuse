"""
Initial Setup: Transactions spreadsheet for COUNT pivot task
Task ID: osworld_calc_pivot_count_invoice_002
Domain: libreoffice_calc

Creates Sheet1 with realistic transaction data and an empty Sheet2 (Summary).
The agent's task is to create a COUNT pivot in Sheet2 grouped by Payment Method.
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_pivot_count_invoice_002'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Transactions ---
    ws1 = wb.active
    ws1.title = 'Transactions'

    headers = ['Transaction ID', 'Date', 'Payment Method', 'Amount', 'Merchant']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # Realistic transaction data
    # Payment method counts:
    #   Credit Card  -> 6
    #   Debit Card   -> 5
    #   PayPal       -> 4
    #   Cash         -> 3
    #   Bank Transfer-> 2
    # Total: 20 transactions
    data = [
        ('TXN-001', '2025-01-03', 'Credit Card',   128.50, 'Amazon'),
        ('TXN-002', '2025-01-05', 'Debit Card',     45.20, 'Whole Foods Market'),
        ('TXN-003', '2025-01-07', 'PayPal',          89.99, 'eBay'),
        ('TXN-004', '2025-01-10', 'Cash',            32.00, 'Corner Bakery'),
        ('TXN-005', '2025-01-12', 'Credit Card',    215.00, 'Best Buy'),
        ('TXN-006', '2025-01-14', 'Bank Transfer', 1200.00, 'Landlord Rent'),
        ('TXN-007', '2025-01-16', 'Debit Card',     67.80, 'Target'),
        ('TXN-008', '2025-01-18', 'PayPal',          14.99, 'Spotify'),
        ('TXN-009', '2025-01-20', 'Cash',            18.50, 'City Diner'),
        ('TXN-010', '2025-01-22', 'Credit Card',    340.00, 'Apple Store'),
        ('TXN-011', '2025-01-24', 'Debit Card',     95.40, 'Costco'),
        ('TXN-012', '2025-01-26', 'PayPal',          29.95, 'Etsy'),
        ('TXN-013', '2025-01-28', 'Cash',            55.00, 'Farmers Market'),
        ('TXN-014', '2025-01-30', 'Credit Card',    178.25, 'Nike'),
        ('TXN-015', '2025-02-02', 'Bank Transfer',  500.00, 'Savings Transfer'),
        ('TXN-016', '2025-02-04', 'Debit Card',     112.60, 'Home Depot'),
        ('TXN-017', '2025-02-06', 'PayPal',          49.00, 'Udemy'),
        ('TXN-018', '2025-02-08', 'Credit Card',    265.75, 'Delta Airlines'),
        ('TXN-019', '2025-02-10', 'Debit Card',      38.90, 'CVS Pharmacy'),
        ('TXN-020', '2025-02-12', 'Credit Card',     92.30, 'Cheesecake Factory'),
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Adjust column widths for readability
    ws1.column_dimensions['A'].width = 14
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 16
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 22

    # --- Sheet 2: Summary (intentionally empty — agent fills this in) ---
    ws2 = wb.create_sheet('Summary')
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
