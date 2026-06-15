"""
Reward Script: Create COUNT pivot showing transactions per payment method in Sheet2 (Summary)
Task ID: osworld_calc_pivot_count_invoice_002
Domain: libreoffice_calc
Scoring:
  - Component 1: Summary sheet has headers "Payment Method" and "Count" (0.25 pts)
  - Component 2: Summary sheet has all 5 expected payment method names (0.35 pts)
  - Component 3: Summary sheet has correct count values per payment method (0.25 pts)
  - Component 4: Summary sheet has a Total row with value 20 (0.15 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_pivot_count_invoice_002'

# Ground truth: expected pivot data from task context
EXPECTED_COUNTS = {
    'Bank Transfer': 2,
    'Cash': 3,
    'Credit Card': 6,
    'Debit Card': 5,
    'PayPal': 4,
}
EXPECTED_TOTAL = 20


def build_pivot_dict(all_rows, col_a_idx, col_b_idx):
    """
    Build a dict mapping payment method -> count from the non-header data rows.
    Excludes any 'total/totals/grand total' rows.
    Returns dict of {str: value}.
    """
    pivot = {}
    for row in all_rows[1:]:  # skip header
        if len(row) <= col_a_idx:
            continue
        label = row[col_a_idx]
        if label is None:
            continue
        label_str = str(label).strip()
        if label_str.lower() in ('total', 'totals', 'grand total'):
            continue
        count_val = row[col_b_idx] if len(row) > col_b_idx else None
        pivot[label_str] = count_val
    return pivot


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Task: Create a COUNT pivot in Summary (Sheet2) showing how many transactions
    occurred per payment method, with column headers and correct totals.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Identify the "Sheet2" / Summary sheet (second sheet in workbook)
    # Task says "Sheet2" but the actual name in the workbook is "Summary"
    summary_sheet = None
    sheet_name_used = None

    if 'Summary' in wb.sheetnames:
        summary_sheet = wb['Summary']
        sheet_name_used = 'Summary'
    elif 'Sheet2' in wb.sheetnames:
        summary_sheet = wb['Sheet2']
        sheet_name_used = 'Sheet2'
    elif len(wb.sheetnames) >= 2:
        sheet_name_used = wb.sheetnames[1]
        summary_sheet = wb[sheet_name_used]

    if summary_sheet is None:
        print("FAIL: No Summary/Sheet2 found in workbook")
        print(f"REWARD: {total_score}")
        return total_score

    print(f"INFO: Using sheet '{sheet_name_used}' as the pivot output sheet")

    # Collect all non-empty rows from Summary sheet
    all_rows = []
    for row in summary_sheet.iter_rows(
            min_row=1, max_row=summary_sheet.max_row,
            min_col=1, max_col=summary_sheet.max_column,
            values_only=True):
        if any(v is not None for v in row):
            all_rows.append(row)

    if not all_rows:
        print("FAIL: Summary sheet is empty — no pivot data found")
        print(f"REWARD: {total_score}")
        return total_score

    # Determine column indices for "Payment Method" and "Count" from header
    col_a_idx = 0  # Payment Method column (default)
    col_b_idx = 1  # Count column (default)
    if all_rows:
        hdr = [str(v).strip().lower() if v is not None else '' for v in all_rows[0]]
        for i, h in enumerate(hdr):
            if 'payment' in h:
                col_a_idx = i
            elif 'count' in h or 'transaction' in h:
                col_b_idx = i

    # Build pivot data dict (excludes header and total rows)
    pivot_data = build_pivot_dict(all_rows, col_a_idx, col_b_idx)

    # -------------------------------------------------------------------
    # Component 1: Headers "Payment Method" and "Count" present (0.25 pts)
    # -------------------------------------------------------------------
    try:
        header_row = all_rows[0] if all_rows else ()
        header_texts = [str(v).strip().lower() if v is not None else '' for v in header_row]

        has_payment_header = any('payment' in h for h in header_texts)
        has_count_header = any('count' in h for h in header_texts)

        if has_payment_header and has_count_header:
            print(f"PASS: Component 1 — Headers found: {list(header_row)} (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — Expected 'Payment Method' and 'Count' headers, "
                  f"found: {list(header_row)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------
    # Component 2: All 5 payment method names present (0.35 pts)
    # -------------------------------------------------------------------
    try:
        expected_methods = set(EXPECTED_COUNTS.keys())
        found_methods = set(pivot_data.keys())
        missing = expected_methods - found_methods

        if not missing:
            print(f"PASS: Component 2 — All 5 payment methods present: "
                  f"{sorted(found_methods)} (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 2 — Missing payment methods: {sorted(missing)}, "
                  f"Found: {sorted(found_methods)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------
    # Component 3: Correct count values per payment method (0.25 pts)
    # -------------------------------------------------------------------
    try:
        count_errors = []
        for method, expected_count in EXPECTED_COUNTS.items():
            actual = pivot_data.get(method)
            if actual is None:
                count_errors.append(f"{method}: missing")
            else:
                try:
                    actual_int = int(float(str(actual)))
                    if actual_int != expected_count:
                        count_errors.append(f"{method}: expected {expected_count}, got {actual_int}")
                except (ValueError, TypeError):
                    count_errors.append(f"{method}: non-numeric value {actual!r}")

        if not count_errors:
            print(f"PASS: Component 3 — All count values correct "
                  f"(BankTransfer=2, Cash=3, CreditCard=6, DebitCard=5, PayPal=4) (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 3 — Incorrect counts: {count_errors}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------
    # Component 4: Total row present with value 20 (0.15 pts)
    # -------------------------------------------------------------------
    try:
        total_label_found = False
        total_count_found = False

        for row in all_rows[1:]:  # skip header
            if len(row) <= col_a_idx:
                continue
            label = row[col_a_idx]
            if label is None:
                continue
            label_str = str(label).strip().lower()
            if label_str in ('total', 'totals', 'grand total'):
                total_label_found = True
                count_val = row[col_b_idx] if len(row) > col_b_idx else None
                if count_val is not None:
                    try:
                        if int(float(str(count_val))) == EXPECTED_TOTAL:
                            total_count_found = True
                        else:
                            print(f"FAIL: Component 4 — Total row found but value is "
                                  f"{count_val}, expected {EXPECTED_TOTAL}")
                    except (ValueError, TypeError):
                        print(f"FAIL: Component 4 — Total row has non-numeric value: {count_val!r}")
                break

        if total_label_found and total_count_found:
            print(f"PASS: Component 4 — Total row found with correct value "
                  f"{EXPECTED_TOTAL} (0.15 pts)")
            total_score += 0.15
        elif not total_label_found:
            print(f"FAIL: Component 4 — No 'Total' row found in Summary sheet")
        # else: already printed specific fail message above
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -------------------------------------------------------------------
    # Final score
    # -------------------------------------------------------------------
    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in a given env
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
