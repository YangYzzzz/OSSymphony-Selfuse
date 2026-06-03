"""
Reward Script: Format HR employee roster as a professional report
Task ID: calc_gsd_029
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20): Title row — merged A1:I1, "Employee Compensation Report 2024", 14pt bold
  Component 2 (0.30): Subtotal rows — 5 department summary rows with SUM formulas, bold, gray bg
  Component 3 (0.20): Currency formatting — Salary/Bonus/Total Comp columns use USD format
  Component 4 (0.15): Conditional formatting — 2 rules on Salary column (>100k green, 70k-100k yellow)
  Component 5 (0.15): Freeze panes — top 2 rows frozen (A3)
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_029'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Roster' not in wb.sheetnames:
        print("FAIL: 'Roster' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Roster']

    # =========================================================================
    # Component 1: Title row — merged A1:I1 with title text, 14pt bold (0.20)
    # This FAILS on initial (no merge, no title row, row 1 has headers)
    # =========================================================================
    try:
        a1 = ws['A1']
        title_val = a1.value
        comp1_score = 0.0

        # Check merged range A1:I1
        has_merge = False
        for mr in ws.merged_cells.ranges:
            mr_str = str(mr).upper()
            if 'A1' in mr_str and 'I1' in mr_str:
                has_merge = True
                break

        # Check title text contains key words
        has_title = (title_val is not None and
                     'employee' in str(title_val).lower() and
                     'compensation' in str(title_val).lower() and
                     'report' in str(title_val).lower() and
                     '2024' in str(title_val))

        # Check font: 14pt bold
        is_bold = a1.font.bold is True
        is_14pt = (a1.font.size is not None and abs(float(a1.font.size) - 14.0) < 0.5)

        if has_merge and has_title:
            comp1_score += 0.10
        if has_title and is_bold and is_14pt:
            comp1_score += 0.10

        if comp1_score > 0:
            print(f"PASS: Component 1 — Title row (merge={has_merge}, title={has_title}, bold={is_bold}, 14pt={is_14pt}) ({comp1_score} pts)")
        else:
            print(f"FAIL: Component 1 — Title row (merge={has_merge}, title='{title_val}', bold={is_bold}, size={a1.font.size})")
        total_score += comp1_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Subtotal rows — 5 dept summary rows with SUM formulas,
    #   bold text, gray background (0.30)
    # This FAILS on initial (no subtotal rows, only 81 rows with no "Summary")
    # =========================================================================
    try:
        comp2_score = 0.0
        expected_depts = ['Engineering', 'Marketing', 'Sales', 'Operations', 'Finance']
        found_subtotals = 0

        for row_idx in range(1, ws.max_row + 1):
            cell_a = ws.cell(row=row_idx, column=1)
            val = cell_a.value
            if val is None or not isinstance(val, str):
                continue
            if 'Summary' not in val:
                continue

            # Check if it matches one of the expected departments
            dept_match = any(dept.lower() in val.lower() for dept in expected_depts)
            if not dept_match:
                continue

            # Check bold
            is_bold = cell_a.font.bold is True

            # Check gray background
            has_gray_bg = False
            try:
                fg_rgb = cell_a.fill.fgColor.rgb
                if fg_rgb and isinstance(fg_rgb, str):
                    # Accept various gray shades (D9D9D9, C0C0C0, BFBFBF, etc.)
                    rgb_hex = fg_rgb[-6:]  # last 6 chars (strip alpha)
                    r, g, b = int(rgb_hex[0:2], 16), int(rgb_hex[2:4], 16), int(rgb_hex[4:6], 16)
                    # Gray: R≈G≈B and reasonably light
                    if abs(r - g) < 30 and abs(g - b) < 30 and r > 150:
                        has_gray_bg = True
            except Exception:
                pass

            # Check SUM formulas in cols E, F, G (Salary, Bonus, Total Comp)
            has_sum_formulas = 0
            for col in [5, 6, 7]:  # E, F, G
                cell_val = ws.cell(row=row_idx, column=col).value
                if cell_val and isinstance(cell_val, str) and '=SUM(' in cell_val.upper():
                    has_sum_formulas += 1

            if is_bold and has_sum_formulas >= 3:
                found_subtotals += 1

        # Award 0.06 per correctly found subtotal row (max 5 * 0.06 = 0.30)
        comp2_score = min(found_subtotals * 0.06, 0.30)

        if found_subtotals > 0:
            print(f"PASS: Component 2 — Found {found_subtotals}/5 subtotal rows ({comp2_score} pts)")
        else:
            print(f"FAIL: Component 2 — No subtotal rows found")
        total_score += comp2_score
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Currency formatting on Salary/Bonus/Total Comp columns (0.20)
    # This FAILS on initial (all cells have 'General' number format)
    # =========================================================================
    try:
        comp3_score = 0.0
        # Find which row has data (skip title row if present)
        # Check a data row (not header, not subtotal) for currency format
        # We'll scan for the first data row with a numeric salary
        currency_cols_formatted = 0

        # Find a data row — look for a row with a numeric value in column 5 (Salary)
        sample_row = None
        for row_idx in range(2, min(ws.max_row + 1, 30)):
            val = ws.cell(row=row_idx, column=5).value
            if val is not None and isinstance(val, (int, float)):
                sample_row = row_idx
                break

        if sample_row:
            for col in [5, 6, 7]:  # E=Salary, F=Bonus, G=Total Comp
                nf = ws.cell(row=sample_row, column=col).number_format
                # Accept formats containing $ sign (e.g. $#,##0.00, $#,##0)
                if nf and '$' in str(nf):
                    currency_cols_formatted += 1

            # Score: need all 3 columns formatted
            if currency_cols_formatted == 3:
                comp3_score = 0.20
            elif currency_cols_formatted >= 1:
                comp3_score = round(currency_cols_formatted * 0.067, 2)

            if comp3_score > 0:
                print(f"PASS: Component 3 — {currency_cols_formatted}/3 currency columns formatted ({comp3_score} pts)")
            else:
                print(f"FAIL: Component 3 — No currency formatting found (sample row {sample_row}, E nf={ws.cell(row=sample_row, column=5).number_format})")
        else:
            print(f"FAIL: Component 3 — Could not find a data row with numeric salary")
        total_score += comp3_score
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: Conditional formatting on Salary column (0.15)
    #   Rule 1: >100000 green fill
    #   Rule 2: 70000-100000 yellow fill
    # This FAILS on initial (0 conditional formatting rules)
    # =========================================================================
    try:
        comp4_score = 0.0
        has_gt_rule = False
        has_between_rule = False

        for cf in ws.conditional_formatting:
            cf_range = str(cf).upper()
            # Check if range covers salary column (column E)
            if 'E' not in cf_range:
                continue

            for rule in cf.rules:
                rule_type = getattr(rule, 'type', '')
                rule_op = getattr(rule, 'operator', '')
                rule_formula = getattr(rule, 'formula', [])

                # Check for greaterThan 100000 rule
                if rule_op == 'greaterThan' and rule_formula:
                    try:
                        threshold = float(rule_formula[0])
                        if threshold == 100000:
                            has_gt_rule = True
                    except (ValueError, IndexError):
                        pass

                # Check for between 70000 and 100000 rule
                if rule_op == 'between' and rule_formula:
                    try:
                        if len(rule_formula) >= 2:
                            low = float(rule_formula[0])
                            high = float(rule_formula[1])
                            if low == 70000 and high == 100000:
                                has_between_rule = True
                    except (ValueError, IndexError):
                        pass

        if has_gt_rule:
            comp4_score += 0.075
        if has_between_rule:
            comp4_score += 0.075

        if comp4_score > 0:
            print(f"PASS: Component 4 — Conditional formatting (>100k={has_gt_rule}, 70k-100k={has_between_rule}) ({comp4_score} pts)")
        else:
            print(f"FAIL: Component 4 — No conditional formatting rules found on Salary column")
        total_score += comp4_score
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # =========================================================================
    # Component 5: Freeze panes — top 2 rows frozen, i.e. freeze_panes == "A3" (0.15)
    # This FAILS on initial (freeze_panes is None)
    # =========================================================================
    try:
        comp5_score = 0.0
        fp = ws.freeze_panes

        if fp is not None and str(fp).upper() == 'A3':
            comp5_score = 0.15
            print(f"PASS: Component 5 — Freeze panes at A3 ({comp5_score} pts)")
        else:
            print(f"FAIL: Component 5 — Freeze panes is '{fp}', expected 'A3'")
        total_score += comp5_score
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved LibreOffice state before verification
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(1.0)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
