"""
Initial Setup: Format HR employee roster as professional report
Task ID: calc_gsd_029
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_029'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Roster'

    # Headers in row 1
    headers = ['Emp ID', 'Name', 'Department', 'Title', 'Salary', 'Bonus', 'Total Comp', 'Hire Date', 'Performance']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Employee data - 80 records sorted by department
    # Engineering: 20 employees (rows 2-21)
    engineering = [
        ['E001', 'Sarah Chen', 'Engineering', 'Senior Software Engineer', 125000, 18750, 143750, '2021-03-15', 'Exceeds'],
        ['E002', 'Marcus Johnson', 'Engineering', 'Staff Engineer', 145000, 21750, 166750, '2019-08-22', 'Exceeds'],
        ['E003', 'Priya Patel', 'Engineering', 'Software Engineer II', 95000, 9500, 104500, '2022-06-01', 'Meets'],
        ['E004', 'James Wilson', 'Engineering', 'DevOps Lead', 118000, 17700, 135700, '2020-01-10', 'Exceeds'],
        ['E005', 'Mei Lin', 'Engineering', 'Frontend Developer', 92000, 9200, 101200, '2022-11-15', 'Meets'],
        ['E006', 'David Kim', 'Engineering', 'Backend Developer', 98000, 14700, 112700, '2021-07-20', 'Meets'],
        ['E007', 'Anna Rodriguez', 'Engineering', 'QA Lead', 105000, 15750, 120750, '2020-04-05', 'Exceeds'],
        ['E008', 'Robert Taylor', 'Engineering', 'Software Engineer I', 78000, 7800, 85800, '2023-09-01', 'Meets'],
        ['E009', 'Yuki Tanaka', 'Engineering', 'Data Engineer', 112000, 16800, 128800, '2021-02-14', 'Exceeds'],
        ['E010', 'Michael Brown', 'Engineering', 'Security Engineer', 120000, 18000, 138000, '2020-06-30', 'Meets'],
        ['E011', 'Fatima Al-Hassan', 'Engineering', 'ML Engineer', 135000, 20250, 155250, '2021-01-18', 'Exceeds'],
        ['E012', 'Chris Martinez', 'Engineering', 'Software Engineer II', 96000, 9600, 105600, '2022-03-22', 'Meets'],
        ['E013', 'Lena Johansson', 'Engineering', 'Platform Engineer', 110000, 16500, 126500, '2020-09-12', 'Meets'],
        ['E014', 'Ahmed Khalil', 'Engineering', 'Software Engineer I', 82000, 8200, 90200, '2023-05-15', 'Developing'],
        ['E015', 'Sophie Turner', 'Engineering', 'Tech Lead', 140000, 21000, 161000, '2019-04-08', 'Exceeds'],
        ['E016', 'Ryan O\'Connor', 'Engineering', 'Software Engineer II', 97000, 9700, 106700, '2022-08-20', 'Meets'],
        ['E017', 'Nadia Petrova', 'Engineering', 'Site Reliability Engineer', 115000, 17250, 132250, '2021-06-01', 'Exceeds'],
        ['E018', 'Tom Zhang', 'Engineering', 'Software Engineer I', 80000, 8000, 88000, '2023-07-10', 'Meets'],
        ['E019', 'Isabella Garcia', 'Engineering', 'Cloud Architect', 150000, 22500, 172500, '2018-11-25', 'Exceeds'],
        ['E020', 'Kevin Park', 'Engineering', 'Software Engineer II', 94000, 9400, 103400, '2022-10-01', 'Meets'],
    ]

    # Marketing: 15 employees (rows 22-36)
    marketing = [
        ['M001', 'Laura White', 'Marketing', 'VP of Marketing', 155000, 31000, 186000, '2018-05-20', 'Exceeds'],
        ['M002', 'Daniel Harris', 'Marketing', 'Content Director', 105000, 15750, 120750, '2020-02-15', 'Meets'],
        ['M003', 'Rachel Green', 'Marketing', 'SEO Manager', 88000, 8800, 96800, '2021-09-01', 'Meets'],
        ['M004', 'Aiden Murphy', 'Marketing', 'Brand Strategist', 92000, 9200, 101200, '2021-04-12', 'Meets'],
        ['M005', 'Chloe Adams', 'Marketing', 'Digital Marketing Lead', 98000, 14700, 112700, '2020-07-25', 'Exceeds'],
        ['M006', 'Nathan Brooks', 'Marketing', 'Marketing Analyst', 75000, 7500, 82500, '2022-12-01', 'Meets'],
        ['M007', 'Olivia Carter', 'Marketing', 'Social Media Manager', 72000, 7200, 79200, '2023-01-15', 'Meets'],
        ['M008', 'Ethan Foster', 'Marketing', 'Product Marketing Manager', 95000, 14250, 109250, '2021-03-08', 'Meets'],
        ['M009', 'Hannah Davis', 'Marketing', 'Email Marketing Specialist', 68000, 6800, 74800, '2023-04-20', 'Developing'],
        ['M010', 'Lucas Mitchell', 'Marketing', 'Graphic Designer', 70000, 7000, 77000, '2022-06-15', 'Meets'],
        ['M011', 'Zoe Bennett', 'Marketing', 'PR Manager', 90000, 9000, 99000, '2021-08-10', 'Meets'],
        ['M012', 'Jack Sullivan', 'Marketing', 'Marketing Coordinator', 62000, 6200, 68200, '2023-06-01', 'Developing'],
        ['M013', 'Mia Thompson', 'Marketing', 'Campaign Manager', 85000, 12750, 97750, '2021-11-22', 'Meets'],
        ['M014', 'Owen Reed', 'Marketing', 'Market Research Analyst', 78000, 7800, 85800, '2022-02-28', 'Meets'],
        ['M015', 'Ava Phillips', 'Marketing', 'Growth Hacker', 82000, 8200, 90200, '2022-09-14', 'Meets'],
    ]

    # Sales: 20 employees (rows 37-56)
    sales = [
        ['S001', 'Brian Cooper', 'Sales', 'VP of Sales', 160000, 40000, 200000, '2017-10-05', 'Exceeds'],
        ['S002', 'Lisa Morgan', 'Sales', 'Regional Sales Director', 130000, 26000, 156000, '2019-01-20', 'Exceeds'],
        ['S003', 'Jake Patterson', 'Sales', 'Account Executive', 85000, 17000, 102000, '2021-05-15', 'Meets'],
        ['S004', 'Samantha Lee', 'Sales', 'Sales Manager', 110000, 22000, 132000, '2020-03-10', 'Exceeds'],
        ['S005', 'Derek Howard', 'Sales', 'Business Development Rep', 65000, 9750, 74750, '2023-02-01', 'Developing'],
        ['S006', 'Emily Richards', 'Sales', 'Account Executive', 88000, 17600, 105600, '2021-08-22', 'Meets'],
        ['S007', 'Carlos Mendez', 'Sales', 'Sales Engineer', 105000, 15750, 120750, '2020-11-15', 'Meets'],
        ['S008', 'Grace Hall', 'Sales', 'Inside Sales Rep', 60000, 9000, 69000, '2023-07-01', 'Developing'],
        ['S009', 'Tyler Nelson', 'Sales', 'Enterprise Account Manager', 120000, 24000, 144000, '2019-06-18', 'Exceeds'],
        ['S010', 'Victoria Chang', 'Sales', 'Channel Sales Manager', 100000, 20000, 120000, '2020-09-25', 'Meets'],
        ['S011', 'Blake Watson', 'Sales', 'Account Executive', 82000, 16400, 98400, '2022-01-10', 'Meets'],
        ['S012', 'Natalie Cruz', 'Sales', 'Sales Operations Analyst', 75000, 7500, 82500, '2022-04-15', 'Meets'],
        ['S013', 'Austin Perry', 'Sales', 'Business Development Rep', 63000, 9450, 72450, '2023-03-20', 'Developing'],
        ['S014', 'Diana Ross', 'Sales', 'Key Account Manager', 115000, 23000, 138000, '2019-12-01', 'Exceeds'],
        ['S015', 'Jordan Bell', 'Sales', 'Sales Trainer', 78000, 7800, 85800, '2021-10-08', 'Meets'],
        ['S016', 'Sophia Rivera', 'Sales', 'Account Executive', 86000, 17200, 103200, '2021-07-14', 'Meets'],
        ['S017', 'Evan Price', 'Sales', 'Sales Manager', 108000, 21600, 129600, '2020-05-22', 'Meets'],
        ['S018', 'Taylor Scott', 'Sales', 'Inside Sales Rep', 62000, 9300, 71300, '2023-01-25', 'Developing'],
        ['S019', 'Nicole Simmons', 'Sales', 'Partnership Manager', 95000, 19000, 114000, '2021-02-10', 'Meets'],
        ['S020', 'Brandon Young', 'Sales', 'Account Executive', 84000, 16800, 100800, '2022-05-30', 'Meets'],
    ]

    # Operations: 15 employees (rows 57-71)
    operations = [
        ['O001', 'Patricia Coleman', 'Operations', 'VP of Operations', 148000, 29600, 177600, '2018-03-15', 'Exceeds'],
        ['O002', 'George Barnes', 'Operations', 'Operations Director', 125000, 18750, 143750, '2019-07-22', 'Exceeds'],
        ['O003', 'Sandra Jenkins', 'Operations', 'Supply Chain Manager', 95000, 9500, 104500, '2021-01-10', 'Meets'],
        ['O004', 'Frank Henderson', 'Operations', 'Logistics Coordinator', 68000, 6800, 74800, '2022-08-05', 'Meets'],
        ['O005', 'Catherine Powell', 'Operations', 'Quality Assurance Manager', 92000, 13800, 105800, '2020-11-20', 'Meets'],
        ['O006', 'Raymond Butler', 'Operations', 'Facilities Manager', 78000, 7800, 85800, '2021-06-15', 'Meets'],
        ['O007', 'Janet Simmons', 'Operations', 'Procurement Specialist', 72000, 7200, 79200, '2022-03-10', 'Meets'],
        ['O008', 'Peter Grant', 'Operations', 'Warehouse Supervisor', 65000, 6500, 71500, '2023-01-05', 'Developing'],
        ['O009', 'Karen Russell', 'Operations', 'Process Improvement Lead', 98000, 14700, 112700, '2020-04-18', 'Meets'],
        ['O010', 'Harold Fisher', 'Operations', 'Fleet Manager', 75000, 7500, 82500, '2022-07-20', 'Meets'],
        ['O011', 'Dorothy Stewart', 'Operations', 'Safety Officer', 82000, 8200, 90200, '2021-09-30', 'Meets'],
        ['O012', 'Wayne Palmer', 'Operations', 'Inventory Analyst', 70000, 7000, 77000, '2022-11-12', 'Meets'],
        ['O013', 'Brenda Diaz', 'Operations', 'Customer Service Manager', 88000, 8800, 96800, '2021-05-25', 'Meets'],
        ['O014', 'Keith Woods', 'Operations', 'Shipping Coordinator', 63000, 6300, 69300, '2023-04-01', 'Developing'],
        ['O015', 'Angela Foster', 'Operations', 'Compliance Officer', 96000, 14400, 110400, '2020-08-14', 'Meets'],
    ]

    # Finance: 10 employees (rows 72-81)
    finance = [
        ['F001', 'Richard Clark', 'Finance', 'CFO', 175000, 43750, 218750, '2017-06-01', 'Exceeds'],
        ['F002', 'Susan Wright', 'Finance', 'Finance Director', 135000, 20250, 155250, '2019-02-15', 'Exceeds'],
        ['F003', 'Thomas Hill', 'Finance', 'Senior Accountant', 92000, 9200, 101200, '2021-04-20', 'Meets'],
        ['F004', 'Jessica Moore', 'Finance', 'Financial Analyst', 85000, 8500, 93500, '2021-10-01', 'Meets'],
        ['F005', 'Andrew Turner', 'Finance', 'Tax Manager', 105000, 15750, 120750, '2020-01-15', 'Meets'],
        ['F006', 'Maria Lopez', 'Finance', 'Accounts Payable Lead', 72000, 7200, 79200, '2022-05-10', 'Meets'],
        ['F007', 'Steven Walker', 'Finance', 'Budget Analyst', 78000, 7800, 85800, '2022-09-01', 'Meets'],
        ['F008', 'Nancy King', 'Finance', 'Payroll Manager', 82000, 8200, 90200, '2021-07-15', 'Meets'],
        ['F009', 'Christopher Allen', 'Finance', 'Internal Auditor', 88000, 8800, 96800, '2021-12-05', 'Meets'],
        ['F010', 'Jennifer Scott', 'Finance', 'Controller', 118000, 17700, 135700, '2019-09-20', 'Exceeds'],
    ]

    all_data = engineering + marketing + sales + operations + finance
    for r, row_data in enumerate(all_data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths
    col_widths = {'A': 8, 'B': 25, 'C': 15, 'D': 30, 'E': 12, 'F': 10, 'G': 12, 'H': 12, 'I': 12}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
