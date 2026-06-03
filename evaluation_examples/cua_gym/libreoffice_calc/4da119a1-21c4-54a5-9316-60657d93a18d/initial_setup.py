"""
Initial Setup: Product Inventory Management Spreadsheet
Task ID: calc_grs_006
Domain: libreoffice_calc

Creates an inventory spreadsheet with:
- 15 rows of realistic product data
- Data validation on Category (dropdown with 6+ categories)
- Data validation on Stock Quantity (whole numbers 0-10000)
- Status column is EMPTY (agent must add IF formula)
- NO conditional formatting (agent must add it)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_006'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    # --- Headers ---
    headers = [
        'SKU', 'Product Name', 'Category', 'Supplier',
        'Unit Cost', 'Selling Price', 'Stock Quantity',
        'Reorder Level', 'Status'
    ]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Product Data (15 rows) ---
    data = [
        ['SKU-1001', 'Wireless Bluetooth Headphones', 'Electronics', 'SoundTech Inc.', 24.50, 49.99, 145, 30],
        ['SKU-1002', 'Organic Cotton T-Shirt (M)', 'Clothing', 'GreenWear Co.', 8.75, 22.00, 12, 50],
        ['SKU-1003', 'Dark Roast Coffee Beans (1kg)', 'Food', 'Mountain Brew Ltd.', 6.20, 14.99, 230, 100],
        ['SKU-1004', 'Cordless Power Drill 18V', 'Tools', 'BuildRight Industries', 45.00, 89.95, 8, 15],
        ['SKU-1005', 'Children\'s Science Encyclopedia', 'Books', 'BrightMinds Publishing', 12.30, 29.99, 67, 25],
        ['SKU-1006', 'Wooden Building Blocks Set', 'Toys', 'PlayWell Toys', 9.80, 24.50, 38, 40],
        ['SKU-1007', 'USB-C Charging Cable (2m)', 'Electronics', 'CablePro Solutions', 3.15, 12.99, 520, 200],
        ['SKU-1008', 'Wool Winter Scarf', 'Clothing', 'NordicKnit Textiles', 14.00, 35.00, 55, 30],
        ['SKU-1009', 'Gluten-Free Pasta Variety Pack', 'Food', 'Bella Italia Foods', 4.50, 11.49, 88, 75],
        ['SKU-1010', 'Adjustable Wrench Set (5pc)', 'Tools', 'IronGrip Hardware', 18.25, 42.00, 22, 20],
        ['SKU-1011', 'Modern Art History Paperback', 'Books', 'Horizon Press', 7.90, 18.50, 41, 15],
        ['SKU-1012', 'Remote Control Racing Car', 'Toys', 'SpeedKidz Entertainment', 16.50, 39.99, 5, 10],
        ['SKU-1013', 'Noise Cancelling Earbuds', 'Electronics', 'SoundTech Inc.', 32.00, 69.99, 95, 40],
        ['SKU-1014', 'Stainless Steel Water Bottle', 'Food', 'EcoSip Products', 5.60, 16.99, 310, 150],
        ['SKU-1015', 'Metric Socket Set (40pc)', 'Tools', 'IronGrip Hardware', 28.75, 64.99, 17, 12],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            # Format currency columns
            if c in (5, 6):
                cell.number_format = '$#,##0.00'
            # Format integer columns
            if c in (7, 8):
                cell.number_format = '0'

    # Status column (I) is intentionally left EMPTY - agent must add IF formula

    # --- Column Widths ---
    col_widths = {'A': 12, 'B': 32, 'C': 14, 'D': 24, 'E': 12, 'F': 14, 'G': 16, 'H': 14, 'I': 14}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # --- Data Validation: Category dropdown ---
    dv_category = DataValidation(
        type="list",
        formula1='"Electronics,Clothing,Food,Tools,Books,Toys"',
        allow_blank=True,
        showDropDown=False,  # False = SHOW the dropdown (counterintuitive)
    )
    dv_category.error = "Please select a valid category from the dropdown list."
    dv_category.errorTitle = "Invalid Category"
    dv_category.prompt = "Select a product category"
    dv_category.promptTitle = "Category"
    dv_category.add("C2:C100")
    ws.add_data_validation(dv_category)

    # --- Data Validation: Stock Quantity (whole numbers 0-10000) ---
    dv_stock = DataValidation(
        type="whole",
        operator="between",
        formula1="0",
        formula2="10000",
        allow_blank=True,
        showErrorMessage=True,
    )
    dv_stock.error = "Stock quantity must be a whole number between 0 and 10,000."
    dv_stock.errorTitle = "Invalid Stock Quantity"
    dv_stock.prompt = "Enter stock quantity (0-10000)"
    dv_stock.promptTitle = "Stock Quantity"
    dv_stock.add("G2:G100")
    ws.add_data_validation(dv_stock)

    # --- Freeze header row ---
    ws.freeze_panes = "A2"

    # --- Auto filter ---
    ws.auto_filter.ref = "A1:I16"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
