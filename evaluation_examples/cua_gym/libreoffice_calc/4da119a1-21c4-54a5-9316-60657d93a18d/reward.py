"""
Reward Script: Product inventory management spreadsheet with IF formulas and conditional formatting
Task ID: calc_grs_006
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): Status column has IF formulas in all data rows
  Component 2 (0.25): IF formula logic is correct (Reorder Now / Low Stock / In Stock)
  Component 3 (0.20): Conditional formatting rule exists targeting Reorder Now status
  Component 4 (0.20): Conditional formatting uses red fill
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_006'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Find the inventory sheet (accept various names)
    ws = None
    for name in wb.sheetnames:
        if name.lower() in ('inventory', 'sheet1', 'sheet'):
            ws = wb[name]
            break
    if ws is None:
        ws = wb.worksheets[0]
    print(f"Using sheet: {ws.title}")

    # Determine data row range (find last row with data in column A)
    max_data_row = 1
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value is not None:
            max_data_row = r
    print(f"Data rows: 2 to {max_data_row}")

    if max_data_row < 2:
        print("FAIL: No data rows found")
        print("REWARD: 0.0")
        return 0.0

    # Find Status column (column I = 9, but search by header name for robustness)
    status_col = None
    stock_col = None
    reorder_col = None
    for c in range(1, ws.max_column + 1):
        hdr = str(ws.cell(1, c).value or "").strip().lower()
        if "status" in hdr:
            status_col = c
        if "stock" in hdr and "quantity" in hdr:
            stock_col = c
        if "reorder" in hdr and "level" in hdr:
            reorder_col = c

    if status_col is None:
        print("FAIL: No 'Status' column found in headers")
        print("REWARD: 0.0")
        return 0.0

    print(f"Status col: {status_col}, Stock Qty col: {stock_col}, Reorder Level col: {reorder_col}")

    # =========================================================================
    # Component 1: Status column has IF formulas in all data rows (0.35 pts)
    # This is the primary task change: Status was empty (None) in initial_env
    # =========================================================================
    try:
        rows_with_formula = 0
        total_data_rows = max_data_row - 1  # rows 2..max_data_row
        for r in range(2, max_data_row + 1):
            val = ws.cell(r, status_col).value
            if val is not None and isinstance(val, str) and val.strip().upper().startswith("=IF"):
                rows_with_formula += 1

        formula_ratio = rows_with_formula / total_data_rows if total_data_rows > 0 else 0
        if formula_ratio >= 0.8:
            print(f"PASS: Component 1 — {rows_with_formula}/{total_data_rows} data rows have IF formulas (0.35 pts)")
            total_score += 0.35
        elif formula_ratio > 0:
            partial = round(0.35 * formula_ratio, 2)
            print(f"PARTIAL: Component 1 — {rows_with_formula}/{total_data_rows} data rows have IF formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No IF formulas found in Status column")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: IF formula logic is correct (0.25 pts)
    # Must check: G<row><H<row> -> "Reorder Now", G<row><H<row>*1.2 -> "Low Stock", else "In Stock"
    # =========================================================================
    try:
        correct_formulas = 0
        for r in range(2, max_data_row + 1):
            val = ws.cell(r, status_col).value
            if val is None or not isinstance(val, str):
                continue
            formula_upper = val.upper().replace(" ", "")
            # Check the formula contains the key elements:
            # - References to stock qty col and reorder level col for this row
            # - "Reorder Now" string
            # - "Low Stock" string
            # - "In Stock" string
            # - Multiplication by 1.2 for the low stock threshold
            has_reorder_now = '"REORDERNOW"' in formula_upper or '"REORDER NOW"' in formula_upper.replace(" ", "X")
            has_low_stock = '"LOWSTOCK"' in formula_upper or '"LOW STOCK"' in formula_upper.replace(" ", "X")
            has_in_stock = '"INSTOCK"' in formula_upper or '"IN STOCK"' in formula_upper.replace(" ", "X")

            # Check for Reorder Now and Low Stock and In Stock in the original (not uppercased-nospace) formula
            orig_nospace = val.replace(" ", "")
            has_reorder_now = '"ReorderNow"' in orig_nospace or '"REORDERNOW"' in orig_nospace.upper().replace(" ", "")
            has_low_stock = '"LowStock"' in orig_nospace or '"LOWSTOCK"' in orig_nospace.upper().replace(" ", "")
            has_in_stock = '"InStock"' in orig_nospace or '"INSTOCK"' in orig_nospace.upper().replace(" ", "")

            # Actually let's just check the raw value for the text strings
            has_reorder_now = "Reorder Now" in val or "reorder now" in val.lower()
            has_low_stock = "Low Stock" in val or "low stock" in val.lower()
            has_in_stock = "In Stock" in val or "in stock" in val.lower()

            # Check for the 1.2 multiplier (20% above threshold)
            has_multiplier = "1.2" in val or "*1.2" in val.replace(" ", "")

            if has_reorder_now and has_low_stock and has_in_stock and has_multiplier:
                correct_formulas += 1

        formula_correct_ratio = correct_formulas / total_data_rows if total_data_rows > 0 else 0
        if formula_correct_ratio >= 0.8:
            print(f"PASS: Component 2 — {correct_formulas}/{total_data_rows} formulas have correct logic (0.25 pts)")
            total_score += 0.25
        elif formula_correct_ratio > 0:
            partial = round(0.25 * formula_correct_ratio, 2)
            print(f"PARTIAL: Component 2 — {correct_formulas}/{total_data_rows} correct formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No formulas with correct Reorder Now/Low Stock/In Stock + 1.2 multiplier logic")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Conditional formatting rule exists for Reorder Now (0.20 pts)
    # Initial env has 0 conditional formatting rules; golden has 1+
    # =========================================================================
    try:
        cf_rules = list(ws.conditional_formatting)
        reorder_cf_found = False
        for cf in cf_rules:
            for rule in cf.rules:
                # Check if rule references "Reorder Now" in its formula
                if rule.formula:
                    for f in rule.formula:
                        if "Reorder Now" in f or "reorder now" in f.lower() or "REORDER NOW" in f.upper():
                            reorder_cf_found = True
                            break
                # Also check expression type rules
                if rule.type == "expression" and rule.formula:
                    for f in rule.formula:
                        if "Reorder" in f or "reorder" in f.lower():
                            reorder_cf_found = True
                            break
                if reorder_cf_found:
                    break
            if reorder_cf_found:
                break

        if reorder_cf_found:
            print(f"PASS: Component 3 — Conditional formatting rule found for 'Reorder Now' (0.20 pts)")
            total_score += 0.20
        elif len(cf_rules) > 0:
            # Has some conditional formatting but not specifically for Reorder Now
            print(f"PARTIAL: Component 3 — {len(cf_rules)} CF rules found but none reference 'Reorder Now' (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — No conditional formatting rules found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: Conditional formatting uses red fill (0.20 pts)
    # The task says "highlight the entire row red when Status is 'Reorder Now'"
    # =========================================================================
    try:
        red_fill_found = False
        for cf in ws.conditional_formatting:
            for rule in cf.rules:
                if rule.dxf and rule.dxf.fill:
                    fill = rule.dxf.fill
                    # Check fgColor for red-ish color
                    fg_rgb = None
                    try:
                        fg_rgb = fill.fgColor.rgb if fill.fgColor else None
                    except Exception:
                        pass
                    if fg_rgb:
                        # Extract RGB components (ARGB format: AARRGGBB)
                        rgb_str = str(fg_rgb)
                        if len(rgb_str) == 8:
                            r_val = int(rgb_str[2:4], 16)
                            g_val = int(rgb_str[4:6], 16)
                            b_val = int(rgb_str[6:8], 16)
                        elif len(rgb_str) == 6:
                            r_val = int(rgb_str[0:2], 16)
                            g_val = int(rgb_str[2:4], 16)
                            b_val = int(rgb_str[4:6], 16)
                        else:
                            continue
                        # Red means high R, low G and B
                        if r_val >= 180 and g_val < 100 and b_val < 100:
                            red_fill_found = True
                            print(f"  Found red fill: RGB({r_val},{g_val},{b_val}) from {rgb_str}")
                            break
            if red_fill_found:
                break

        if red_fill_found:
            print(f"PASS: Component 4 — Conditional formatting uses red fill (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 4 — No red fill found in conditional formatting rules")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
