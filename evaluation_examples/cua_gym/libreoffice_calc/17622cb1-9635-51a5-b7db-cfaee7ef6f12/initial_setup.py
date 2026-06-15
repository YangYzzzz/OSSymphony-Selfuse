"""
Initial Setup: Multi-year sales data for pivot table and chart creation
Task ID: calc_gcp_048
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_048'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def random_date(start_year, end_year):
    """Generate a random date between start_year-01-01 and end_year-12-31."""
    start = date(start_year, 1, 1)
    end = date(end_year, 12, 31)
    delta = (end - start).days
    return start + timedelta(days=random.randint(0, delta))


def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'MultiYearSales'

    # --- Headers ---
    headers = ['SaleID', 'SaleDate', 'Region', 'Product', 'Quantity', 'Revenue']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Column widths ---
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 14

    # --- Data config ---
    regions = ['Americas', 'EMEA', 'APAC']
    products = [
        'Cloud Compute Pro', 'Data Analytics Suite', 'AI Platform License',
        'Security Shield', 'DevOps Toolkit', 'Storage Optimizer',
        'Network Gateway', 'API Management Hub', 'ML Pipeline Runner',
        'Identity Manager', 'Monitoring Dashboard', 'Database Engine',
        'Edge Computing Node', 'Compliance Tracker', 'Collaboration Suite',
    ]

    # Revenue base ranges by region (to create visible trends)
    region_base = {
        'Americas': (8000, 65000),
        'EMEA': (6000, 55000),
        'APAC': (5000, 50000),
    }

    # Growth multiplier by year (simulate growth trend)
    year_multiplier = {2022: 1.0, 2023: 1.15, 2024: 1.32}

    # --- Generate 800 rows ---
    for i in range(1, 801):
        row = i + 1
        sale_date = random_date(2022, 2024)
        region = random.choice(regions)
        product = random.choice(products)
        quantity = random.randint(1, 50)
        base_lo, base_hi = region_base[region]
        multiplier = year_multiplier[sale_date.year]
        revenue = round(random.uniform(base_lo, base_hi) * multiplier, 2)

        ws.cell(row=row, column=1, value=i)
        date_cell = ws.cell(row=row, column=2, value=sale_date)
        date_cell.number_format = 'yyyy-mm-dd'
        ws.cell(row=row, column=3, value=region)
        ws.cell(row=row, column=4, value=product)
        ws.cell(row=row, column=5, value=quantity)
        rev_cell = ws.cell(row=row, column=6, value=revenue)
        rev_cell.number_format = '$#,##0.00'

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
