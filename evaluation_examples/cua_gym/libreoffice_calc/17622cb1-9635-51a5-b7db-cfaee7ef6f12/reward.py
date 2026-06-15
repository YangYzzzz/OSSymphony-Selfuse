"""
Reward Script: Create pivot table from multi-year sales data with grouped bar chart
Task ID: calc_gcp_048
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20): PivotTable sheet exists in the workbook
  Component 2 (0.30): Pivot table structure — 12 Year-Quarter rows, 3 region columns
  Component 3 (0.20): Revenue values are numeric and non-trivial aggregations
  Component 4 (0.30): Grouped bar chart with 3 series (one per region)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_048'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # ---------------------------------------------------------------
    # Component 1: PivotTable sheet exists (0.20 points)
    # This is the most basic requirement — a new sheet for the pivot.
    # Initial file only has 'MultiYearSales', so this FAILS on initial.
    # ---------------------------------------------------------------
    pivot_sheet = None
    try:
        # Look for a sheet whose name contains "pivot" (case-insensitive)
        # or any sheet that is NOT 'MultiYearSales'
        pivot_candidates = [s for s in wb.sheetnames if s != 'MultiYearSales']
        if len(pivot_candidates) >= 1:
            pivot_sheet = wb[pivot_candidates[0]]
            print(f"PASS: Component 1 — Found pivot sheet '{pivot_candidates[0]}' (0.20 pts)")
            total_score += 0.20
        else:
            print("FAIL: Component 1 — No additional sheet found besides 'MultiYearSales'")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    if pivot_sheet is None:
        # Cannot proceed without a pivot sheet
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    # ---------------------------------------------------------------
    # Component 2: Pivot table structure (0.30 points)
    # Must have: Year-Quarter grouping as rows (12 groups: 2022-Q1 to 2024-Q4)
    # and Region as columns (Americas, EMEA, APAC).
    # Sub-checks:
    #   2a (0.10): Has header row with region names
    #   2b (0.10): Has 12 data rows (one per Year-Quarter combo)
    #   2c (0.10): Row labels cover all 3 years (2022, 2023, 2024) and 4 quarters
    # ---------------------------------------------------------------
    try:
        # 2a: Check header row contains region names
        header_vals = []
        for c in range(1, pivot_sheet.max_column + 1):
            v = pivot_sheet.cell(row=1, column=c).value
            if v is not None:
                header_vals.append(str(v).strip())

        regions_found = set()
        expected_regions = {'Americas', 'EMEA', 'APAC'}
        for h in header_vals:
            for reg in expected_regions:
                if reg.lower() in h.lower():
                    regions_found.add(reg)

        if len(regions_found) >= 3:
            print(f"PASS: Component 2a — Header contains all 3 regions: {regions_found} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2a — Expected 3 regions in header, found: {regions_found} (headers: {header_vals})")

        # 2b: Check for ~12 data rows (Year-Quarter combos)
        # Count rows that have numeric data in at least one region column
        # Find region column indices
        region_cols = {}
        for c in range(1, pivot_sheet.max_column + 1):
            hv = pivot_sheet.cell(row=1, column=c).value
            if hv is not None:
                hv_str = str(hv).strip()
                for reg in expected_regions:
                    if reg.lower() == hv_str.lower():
                        region_cols[reg] = c

        data_row_count = 0
        for r in range(2, pivot_sheet.max_row + 1):
            # Check if any region column in this row has a numeric value
            has_numeric = False
            for reg, col_idx in region_cols.items():
                val = pivot_sheet.cell(row=r, column=col_idx).value
                if isinstance(val, (int, float)):
                    has_numeric = True
                    break
            # Also check if first column has "Grand Total" — skip that
            first_val = pivot_sheet.cell(row=r, column=1).value
            if first_val is not None and 'grand' in str(first_val).lower():
                continue
            if has_numeric:
                data_row_count += 1

        if data_row_count >= 11 and data_row_count <= 13:
            print(f"PASS: Component 2b — Found {data_row_count} data rows (~12 Year-Quarter combos) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2b — Expected ~12 data rows, found: {data_row_count}")

        # 2c: Check that data covers all 3 years and 4 quarters
        years_found = set()
        quarters_found = set()
        for r in range(2, pivot_sheet.max_row + 1):
            for c in range(1, pivot_sheet.max_column + 1):
                val = pivot_sheet.cell(row=r, column=c).value
                if val is not None:
                    val_str = str(val).strip()
                    # Check for year values
                    if val_str in ('2022', '2023', '2024'):
                        years_found.add(int(val_str))
                    elif isinstance(val, (int, float)) and val in (2022, 2023, 2024):
                        years_found.add(int(val))
                    # Check for quarter values
                    if val_str in ('Q1', 'Q2', 'Q3', 'Q4'):
                        quarters_found.add(val_str)
                    # Also check combined format like "2022-Q1"
                    for y in (2022, 2023, 2024):
                        for q in ('Q1', 'Q2', 'Q3', 'Q4'):
                            if f"{y}-{q}" in val_str or f"{y} {q}" in val_str:
                                years_found.add(y)
                                quarters_found.add(q)

        if len(years_found) >= 3 and len(quarters_found) >= 4:
            print(f"PASS: Component 2c — Covers years {years_found} and quarters {quarters_found} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2c — Expected 3 years & 4 quarters, found years={years_found}, quarters={quarters_found}")

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---------------------------------------------------------------
    # Component 3: Revenue values are numeric and non-trivial (0.20 points)
    # The pivot should contain SUM of Revenue values that are large numbers
    # (hundreds of thousands per Year-Quarter-Region).
    # Sub-checks:
    #   3a (0.10): At least 30 numeric revenue cells in the pivot
    #   3b (0.10): Revenue values are in plausible range (> 10000)
    # ---------------------------------------------------------------
    try:
        numeric_count = 0
        large_value_count = 0
        for r in range(2, pivot_sheet.max_row + 1):
            for reg, col_idx in region_cols.items():
                val = pivot_sheet.cell(row=r, column=col_idx).value
                if isinstance(val, (int, float)):
                    numeric_count += 1
                    if val > 10000:
                        large_value_count += 1

        if numeric_count >= 30:
            print(f"PASS: Component 3a — {numeric_count} numeric revenue cells found (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3a — Expected >=30 numeric revenue cells, found {numeric_count}")

        if large_value_count >= 30:
            print(f"PASS: Component 3b — {large_value_count} revenue values > 10000 (plausible aggregations) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3b — Expected >=30 large revenue values, found {large_value_count}")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ---------------------------------------------------------------
    # Component 4: Grouped bar chart (0.30 points)
    # Must have a bar/column chart with clustered grouping and 3 series.
    # Sub-checks:
    #   4a (0.10): Chart exists on the pivot sheet
    #   4b (0.10): Chart is a bar/column type with clustered grouping
    #   4c (0.10): Chart has 3 data series (one per region)
    # ---------------------------------------------------------------
    try:
        charts = pivot_sheet._charts
        if len(charts) >= 1:
            print(f"PASS: Component 4a — Found {len(charts)} chart(s) on pivot sheet (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4a — No charts found on pivot sheet")
            # Check other sheets for charts
            for sn in wb.sheetnames:
                other_ws = wb[sn]
                if len(other_ws._charts) > 0:
                    print(f"  NOTE: Found {len(other_ws._charts)} chart(s) on sheet '{sn}'")
                    charts = other_ws._charts
                    total_score += 0.10
                    print(f"  PASS: Component 4a — Chart found on sheet '{sn}' (0.10 pts)")
                    break

        if len(charts) >= 1:
            chart = charts[0]
            # 4b: Chart type check — should be bar/column chart
            chart_class = chart.__class__.__name__
            is_bar_type = False
            try:
                if chart_class == 'BarChart':
                    is_bar_type = True
                elif hasattr(chart, 'tagname') and 'bar' in chart.tagname.lower():
                    is_bar_type = True
            except:
                pass

            is_grouped = False
            try:
                if hasattr(chart, 'grouping') and chart.grouping in ('clustered', 'stacked', 'percentStacked'):
                    is_grouped = True
            except:
                pass

            if is_bar_type and is_grouped:
                print(f"PASS: Component 4b — Chart is {chart_class} with grouping='{chart.grouping}' (0.10 pts)")
                total_score += 0.10
            elif is_bar_type:
                print(f"PARTIAL: Component 4b — Chart is bar type but grouping check unclear (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 4b — Expected bar/column chart, found {chart_class}")

            # 4c: Series count — should have 3 series (Americas, EMEA, APAC)
            try:
                series_count = len(chart.series)
                if series_count == 3:
                    print(f"PASS: Component 4c — Chart has {series_count} data series (0.10 pts)")
                    total_score += 0.10
                elif series_count >= 2:
                    print(f"PARTIAL: Component 4c — Chart has {series_count} series, expected 3 (0.05 pts)")
                    total_score += 0.05
                else:
                    print(f"FAIL: Component 4c — Chart has {series_count} series, expected 3")
            except Exception as e:
                print(f"ERROR: Component 4c — {e}")

    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
