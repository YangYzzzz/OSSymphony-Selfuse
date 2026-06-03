"""
Initial Setup: Succession planning matrix with Performance vs Potential ratings
Task ID: calc_hr_061
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_061'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Succession ---
    ws = wb.active
    ws.title = 'Succession'

    # Headers
    headers = ['Employee', 'Performance (1-5)', 'Potential (1-5)', '9-Box Category']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Employee data - Performance and Potential scores
    # D column is intentionally EMPTY (task is to add the nested IF formulas)
    data = [
        ['Alice',  5, 5],
        ['Bob',    4, 3],
        ['Carol',  2, 4],
        ['Dan',    3, 3],
        ['Eve',    5, 3],
        ['Frank',  1, 2],
    ]

    data_font = Font(name='Calibri', size=11)
    center_align = Alignment(horizontal='center', vertical='center')

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = thin_border
            if c >= 2:
                cell.alignment = center_align
            else:
                cell.alignment = Alignment(vertical='center')

        # D column cell - empty but bordered (ready for formula)
        d_cell = ws.cell(row=r, column=4)
        d_cell.border = thin_border
        d_cell.alignment = center_align

    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 22

    # Row height for header
    ws.row_dimensions[1].height = 30

    # Freeze header row
    ws.freeze_panes = 'A2'

    # --- Sheet: 9-Box Reference ---
    ws2 = wb.create_sheet('9-Box Reference')
    ws2['A1'] = '9-Box Grid Reference'
    ws2['A1'].font = Font(name='Calibri', size=14, bold=True)

    # Reference grid headers
    ws2['A3'] = 'Performance \\ Potential'
    ws2['B3'] = 'High (4-5)'
    ws2['C3'] = 'Medium (3)'
    ws2['D3'] = 'Low (1-2)'
    for col in range(1, 5):
        ws2.cell(row=3, column=col).font = Font(bold=True)
        ws2.cell(row=3, column=col).alignment = Alignment(horizontal='center')

    # Grid body
    ref_data = [
        ['High (4-5)',  'Star',           'High Performer', 'Specialist'],
        ['Medium (3)',  'High Potential',  'Core Player',    'Average'],
        ['Low (1-2)',   'Enigma',         'Underperformer', 'Risk'],
    ]
    for r, row_data in enumerate(ref_data, 4):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if c == 1:
                cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
