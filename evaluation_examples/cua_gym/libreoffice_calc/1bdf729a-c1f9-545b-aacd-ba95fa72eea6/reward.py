"""
Reward Script: Succession planning matrix with 9-Box classification formulas
Task ID: calc_hr_061
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.3): All D2:D7 cells contain formulas (not empty)
  - Component 2 (0.3): Formulas use nested IF logic referencing B and C columns
  - Component 3 (0.4): Formulas produce correct 9-Box categories for each employee
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_061'

# Ground truth mapping: (Performance, Potential) -> expected category
EXPECTED = {
    2: ('Alice',  5, 5, 'Star'),
    3: ('Bob',    4, 3, 'High Performer'),
    4: ('Carol',  2, 4, 'Enigma'),
    5: ('Dan',    3, 3, 'Core Player'),
    6: ('Eve',    5, 3, 'High Performer'),
    7: ('Frank',  1, 2, 'Risk'),
}


def classify_9box(perf, pot):
    """
    Manually evaluate 9-Box classification logic.
    High Perf (4-5), Med Perf (3), Low Perf (1-2)
    High Pot (4-5), Med Pot (3), Low Pot (1-2)
    """
    if perf >= 4:
        if pot >= 4:
            return 'Star'
        elif pot == 3:
            return 'High Performer'
        else:
            return 'Specialist'
    elif perf == 3:
        if pot >= 4:
            return 'High Potential'
        elif pot == 3:
            return 'Core Player'
        else:
            return 'Average'
    else:  # perf <= 2
        if pot >= 4:
            return 'Enigma'
        elif pot == 3:
            return 'Underperformer'
        else:
            return 'Risk'


def evaluate_formula_result(formula_str, perf_val, pot_val):
    """
    Parse and evaluate a nested IF formula to determine its output
    given performance and potential values. Returns the string result
    or None if parsing fails.
    """
    # Replace cell references with actual values
    # The formula references B<row> for perf and C<row> for pot
    # We substitute all B-column refs with perf_val and C-column refs with pot_val
    expr = formula_str
    if not isinstance(expr, str) or not expr.startswith('='):
        return None

    expr = expr[1:]  # remove leading '='

    # Replace cell references like B2, C2 etc with actual numeric values
    expr = re.sub(r'[Bb]\d+', str(perf_val), expr)
    expr = re.sub(r'[Cc]\d+', str(pot_val), expr)

    # Now evaluate the nested IF expression
    try:
        result = eval_if_expr(expr)
        return result
    except Exception:
        return None


def eval_if_expr(expr):
    """
    Recursively evaluate Excel IF(condition, true_val, false_val) expressions.
    Supports: IF, >=, <=, >, <, =, numeric comparisons, string literals.
    """
    expr = expr.strip()

    # String literal
    if expr.startswith('"') and expr.endswith('"'):
        return expr[1:-1]

    # Numeric literal
    try:
        return float(expr)
    except ValueError:
        pass

    # IF function
    if expr.upper().startswith('IF('):
        # Find the matching closing paren
        inner = extract_if_args(expr)
        if inner is None:
            return None
        condition, true_val, false_val = inner
        cond_result = eval_condition(condition)
        if cond_result:
            return eval_if_expr(true_val)
        else:
            return eval_if_expr(false_val)

    return None


def extract_if_args(expr):
    """Extract the three arguments of IF(...) handling nested IFs."""
    # Remove IF( prefix and trailing )
    if not expr.upper().startswith('IF('):
        return None

    inner = expr[3:]  # remove 'IF('
    # Find matching closing paren
    depth = 1
    end = 0
    for i, ch in enumerate(inner):
        if ch == '(':
            depth += 1
        elif ch == ')':
            depth -= 1
            if depth == 0:
                end = i
                break

    content = inner[:end]

    # Split by commas at depth 0
    parts = []
    current = []
    depth = 0
    for ch in content:
        if ch == '(':
            depth += 1
        elif ch == ')':
            depth -= 1
        elif ch == ',' and depth == 0:
            parts.append(''.join(current).strip())
            current = []
            continue
        current.append(ch)
    parts.append(''.join(current).strip())

    if len(parts) != 3:
        return None
    return parts[0], parts[1], parts[2]


def eval_condition(cond):
    """Evaluate a simple condition like '5>=4' or '3=3'."""
    cond = cond.strip()

    # Handle >=, <=, >, <, =
    for op in ['>=', '<=', '>', '<', '=']:
        # Split on the operator (but not inside strings)
        idx = cond.find(op)
        if idx > 0:
            left = cond[:idx].strip()
            right = cond[idx + len(op):].strip()
            try:
                l = float(left)
                r = float(right)
                if op == '>=':
                    return l >= r
                elif op == '<=':
                    return l <= r
                elif op == '>':
                    return l > r
                elif op == '<':
                    return l < r
                elif op == '=':
                    return l == r
            except ValueError:
                continue

    return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print("CRITICAL: Cannot load file {}: {}".format(file_path, e))
        print("REWARD: 0.0")
        return 0.0

    if 'Succession' not in wb.sheetnames:
        print("CRITICAL: 'Succession' sheet not found. Sheets: {}".format(wb.sheetnames))
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Succession']

    # Component 1: All D2:D7 cells contain formulas (0.3 points)
    # This checks that the agent entered SOMETHING in the D column
    try:
        formula_count = 0
        for row in range(2, 8):
            val = ws.cell(row=row, column=4).value
            if val is not None and isinstance(val, str) and val.startswith('='):
                formula_count += 1

        if formula_count == 6:
            print("PASS: Component 1 - All 6 cells D2:D7 contain formulas (0.3 pts)")
            total_score += 0.3
        elif formula_count > 0:
            partial = round(0.3 * formula_count / 6, 2)
            print("PARTIAL: Component 1 - {}/6 cells have formulas ({} pts)".format(formula_count, partial))
            total_score += partial
        else:
            print("FAIL: Component 1 - No formulas found in D2:D7")
    except Exception as e:
        print("ERROR: Component 1 - {}".format(e))

    # Component 2: Formulas use nested IF logic with B and C references (0.3 points)
    # This checks that the formulas are structurally correct (nested IFs, not just static text)
    try:
        valid_formula_count = 0
        for row in range(2, 8):
            val = ws.cell(row=row, column=4).value
            if val and isinstance(val, str) and val.startswith('='):
                upper_val = val.upper()
                has_if = 'IF(' in upper_val
                has_b_ref = bool(re.search(r'B\d+', val, re.IGNORECASE))
                has_c_ref = bool(re.search(r'C\d+', val, re.IGNORECASE))
                has_nested_if = upper_val.count('IF(') >= 2

                if has_if and has_b_ref and has_c_ref and has_nested_if:
                    valid_formula_count += 1

        if valid_formula_count == 6:
            print("PASS: Component 2 - All 6 formulas use nested IF with B/C refs (0.3 pts)")
            total_score += 0.3
        elif valid_formula_count > 0:
            partial = round(0.3 * valid_formula_count / 6, 2)
            print("PARTIAL: Component 2 - {}/6 formulas are structurally valid ({} pts)".format(valid_formula_count, partial))
            total_score += partial
        else:
            print("FAIL: Component 2 - No valid nested IF formulas found")
    except Exception as e:
        print("ERROR: Component 2 - {}".format(e))

    # Component 3: Formulas produce correct 9-Box categories (0.4 points)
    # We evaluate each formula against the known B/C values and check against ground truth
    try:
        correct_count = 0
        for row, (name, perf, pot, expected_cat) in EXPECTED.items():
            val = ws.cell(row=row, column=4).value

            # Read the actual B and C values from the sheet
            actual_perf = ws.cell(row=row, column=2).value
            actual_pot = ws.cell(row=row, column=3).value

            if actual_perf is None or actual_pot is None:
                print("FAIL: Component 3 row {} - Missing B/C values".format(row))
                continue

            if val and isinstance(val, str) and val.startswith('='):
                # Evaluate the formula
                result = evaluate_formula_result(val, int(actual_perf), int(actual_pot))
                if result == expected_cat:
                    correct_count += 1
                    print("PASS: Component 3 row {} ({}) - Formula evaluates to '{}' (correct)".format(row, name, result))
                else:
                    print("FAIL: Component 3 row {} ({}) - Formula evaluates to '{}', expected '{}'".format(row, name, result, expected_cat))
            elif val and isinstance(val, str) and not val.startswith('='):
                # Plain text value (not a formula but might match)
                if val.strip() == expected_cat:
                    # Give partial: correct value but not a formula
                    correct_count += 0.5
                    print("PARTIAL: Component 3 row {} ({}) - Static text '{}' matches but is not a formula".format(row, name, val))
                else:
                    print("FAIL: Component 3 row {} ({}) - Static text '{}', expected '{}'".format(row, name, val, expected_cat))
            else:
                print("FAIL: Component 3 row {} ({}) - D{} is empty".format(row, name, row))

        if correct_count == 6:
            print("PASS: Component 3 - All 6 categories correct (0.4 pts)")
            total_score += 0.4
        elif correct_count > 0:
            partial = round(0.4 * correct_count / 6, 2)
            print("PARTIAL: Component 3 - {}/6 categories correct ({} pts)".format(correct_count, partial))
            total_score += partial
        else:
            print("FAIL: Component 3 - No correct categories")
    except Exception as e:
        print("ERROR: Component 3 - {}".format(e))

    final_score = min(round(total_score, 2), 1.0)
    print("\nScore: {}/1.0".format(total_score))
    print("REWARD: {}".format(final_score))
    return final_score


# Entry point
file_path = os.path.join(WORKDIR, TASK_ID + '.xlsx')
if not os.path.exists(file_path):
    print("File not found: {}".format(file_path))
    print("REWARD: 0.0")
else:
    verify_task(file_path)
