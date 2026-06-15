"""
Reward Script: Nested IF formula for letter grades in column E
Task ID: calc_gg5_014
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Coverage — E2:E80 all contain formulas (>=95% filled)
  Component 2 (0.3): Structure — formulas are nested IF with correct grade thresholds
  Component 3 (0.3): Row consistency — each formula references its own row's D cell
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_014'


def persist_app_state(domain):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Results' sheet must exist
    if 'Results' not in wb.sheetnames:
        print("FAIL: 'Results' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Results']

    # Component 1: Coverage — E2:E80 contain formulas (0.4 points)
    # Only cells with formula strings (starting with '=') count.
    # This FAILS on initial (all None) and PASSES on golden (all formulas).
    try:
        formula_count = 0
        for r in range(2, 81):
            val = ws.cell(row=r, column=5).value
            if val is not None and isinstance(val, str) and val.startswith('='):
                formula_count += 1

        coverage = formula_count / 79.0
        print(f"Component 1: {formula_count}/79 cells in E2:E80 contain formulas (coverage={coverage:.2%})")

        if coverage >= 0.95:
            print(f"PASS: Component 1 — formula coverage >= 95% (0.4 pts)")
            total_score += 0.4
        elif coverage >= 0.5:
            partial = 0.4 * (coverage - 0.5) / 0.45  # linear from 50% to 95%
            print(f"PARTIAL: Component 1 — coverage {coverage:.0%}, awarding {partial:.2f} pts")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — only {formula_count}/79 cells have formulas")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Correct nested IF structure with grade thresholds (0.3 points)
    # Check that formulas contain the right grade boundaries: >=90 -> A, >=80 -> B, >=70 -> C, >=60 -> D, else F
    try:
        correct_structure = 0
        sampled = 0
        # Check all formula cells
        for r in range(2, 81):
            val = ws.cell(row=r, column=5).value
            if val is None or not isinstance(val, str) or not val.startswith('='):
                continue
            sampled += 1
            formula_upper = val.upper().replace(" ", "")
            # Check all 5 grade thresholds are present
            has_a_90 = ('>=90,"A"' in formula_upper or '>=90,"a"' in formula_upper.lower())
            has_b_80 = ('>=80,"B"' in formula_upper or '>=80,"b"' in formula_upper.lower())
            has_c_70 = ('>=70,"C"' in formula_upper or '>=70,"c"' in formula_upper.lower())
            has_d_60 = ('>=60,"D"' in formula_upper or '>=60,"d"' in formula_upper.lower())
            has_f = ('"F"' in formula_upper or '"f"' in formula_upper.lower())

            if has_a_90 and has_b_80 and has_c_70 and has_d_60 and has_f:
                correct_structure += 1

        if sampled == 0:
            print("FAIL: Component 2 — no formulas found to check structure")
        else:
            ratio = correct_structure / sampled
            print(f"Component 2: {correct_structure}/{sampled} formulas have correct IF structure (ratio={ratio:.2%})")
            if ratio >= 0.95:
                print(f"PASS: Component 2 — correct nested IF structure (0.3 pts)")
                total_score += 0.3
            elif ratio >= 0.5:
                partial = 0.3 * (ratio - 0.5) / 0.45
                print(f"PARTIAL: Component 2 — {ratio:.0%} correct, awarding {partial:.2f} pts")
                total_score += partial
            else:
                print(f"FAIL: Component 2 — only {correct_structure}/{sampled} have correct structure")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Row-consistent D references (0.3 points)
    # Each formula in row r should reference D{r} (not some other row).
    # e.g., E5 should contain "D5" in its formula, not "D2" or "$D$2".
    try:
        consistent = 0
        checked = 0
        for r in range(2, 81):
            val = ws.cell(row=r, column=5).value
            if val is None or not isinstance(val, str) or not val.startswith('='):
                continue
            checked += 1
            formula_upper = val.upper().replace(" ", "")
            expected_ref = f"D{r}"
            # The formula should reference D{r} (the same row), not a fixed row
            if expected_ref.upper() in formula_upper:
                consistent += 1

        if checked == 0:
            print("FAIL: Component 3 — no formulas to check row consistency")
        else:
            ratio = consistent / checked
            print(f"Component 3: {consistent}/{checked} formulas reference correct D row (ratio={ratio:.2%})")
            if ratio >= 0.95:
                print(f"PASS: Component 3 — row-consistent references (0.3 pts)")
                total_score += 0.3
            elif ratio >= 0.5:
                partial = 0.3 * (ratio - 0.5) / 0.45
                print(f"PARTIAL: Component 3 — {ratio:.0%} consistent, awarding {partial:.2f} pts")
                total_score += partial
            else:
                print(f"FAIL: Component 3 — only {consistent}/{checked} reference correct row")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
