"""
Initial Setup: Create exam results spreadsheet with 79 students, scores in D, column E empty.
Task ID: calc_gg5_014
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_014'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    # --- Headers ---
    headers = ["Student ID", "Name", "Subject", "Score", "Grade"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Student data ---
    first_names = [
        "Sarah", "Marcus", "Elena", "James", "Priya", "David", "Mei", "Carlos",
        "Aisha", "Ryan", "Sofia", "Tyler", "Nadia", "Kevin", "Luna", "Omar",
        "Grace", "Andre", "Yuki", "Patrick", "Zara", "Daniel", "Olivia", "Hassan",
        "Emma", "Jamal", "Mia", "Thomas", "Fatima", "Lucas", "Aria", "Nathan",
        "Leila", "Brandon", "Chloe", "Ibrahim", "Hannah", "Victor", "Amara", "Ethan",
        "Sakura", "Diego", "Rebecca", "Jin", "Lily", "Antonio", "Kayla", "Wei",
        "Samantha", "Raj", "Isabelle", "Felix", "Aaliyah", "Mikhail", "Clara",
        "Adrian", "Jade", "Kwame", "Violet", "Sebastian", "Rosa", "Liam",
        "Freya", "Tariq", "Stella", "Mateo", "Bianca", "Noah", "Simone",
        "Alex", "Kira", "Gabriel", "Tanya", "Leo", "Daria", "Oscar", "Vera",
        "Hugo", "Maya",
    ]

    last_names = [
        "Chen", "Johnson", "Kowalski", "Okonkwo", "Patel", "Kim", "Hernandez",
        "Williams", "Nguyen", "Brown", "Ivanov", "Santos", "Anderson", "Tanaka",
        "Garcia", "Ali", "Mueller", "Davis", "Johansson", "Lopez", "Singh",
        "Wilson", "Park", "Thomas", "Martinez", "Lee", "Robinson", "Yamamoto",
        "Taylor", "Abbas", "Schmidt", "Jackson", "Petrov", "Clark", "Nakamura",
        "Rodriguez", "White", "Sato", "Harris", "Khan", "Berg", "Lewis",
        "Suzuki", "Walker", "Morales", "Hall", "Ito", "Young", "Diaz",
        "Allen", "Kimura", "King", "Chandra", "Wright", "Costa", "Scott",
        "Fujita", "Green", "Okafor", "Adams", "Sharma", "Baker", "Torres",
        "Hill", "Watanabe", "Campbell", "Flores", "Mitchell", "Hayashi",
        "Roberts", "Silva", "Carter", "Aoki", "Phillips", "Mendez", "Evans",
        "Turner", "Gupta", "Morris",
    ]

    subjects = [
        "Mathematics", "Physics", "Chemistry", "Biology", "English Literature",
        "Computer Science", "History", "Economics", "Psychology", "Statistics",
        "Philosophy", "Sociology", "Environmental Science",
    ]

    # Generate scores with a realistic distribution
    scores = []
    for _ in range(79):
        # Use a normal distribution centered around 72 with std dev 15
        score = int(random.gauss(72, 15))
        score = max(0, min(100, score))
        scores.append(score)

    for i in range(79):
        row = i + 2
        student_id = f"STU-{2024000 + i + 1:07d}"
        name = f"{first_names[i]} {last_names[i]}"
        subject = subjects[i % len(subjects)]
        score = scores[i]

        ws.cell(row=row, column=1, value=student_id)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=subject)
        ws.cell(row=row, column=4, value=score)
        # Column E (Grade) is intentionally left EMPTY

    # --- Column widths ---
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f"Initial file created: {OUTPUT}")

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print("GUI_READY: launched LibreOffice Calc with DISPLAY=:0")


create_initial()
