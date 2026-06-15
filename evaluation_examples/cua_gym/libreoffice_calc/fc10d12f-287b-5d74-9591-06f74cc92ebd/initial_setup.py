"""
Initial Setup: Apply alternating row shading via conditional formatting
Task ID: calc_gg3_017
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_017'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Roster'

    # --- Headers (Row 1) ---
    headers = ['ID', 'Name', 'Team', 'Role', 'Level', 'Location', 'Phone']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # --- 30 rows of realistic staff data (rows 2-31) ---
    data = [
        ['E001', 'Sarah Chen', 'Engineering', 'Software Engineer', 'Senior', 'San Francisco', '(415) 555-0101'],
        ['E002', 'Marcus Johnson', 'Marketing', 'Campaign Manager', 'Mid', 'New York', '(212) 555-0102'],
        ['E003', 'Priya Patel', 'Engineering', 'DevOps Lead', 'Senior', 'Austin', '(512) 555-0103'],
        ['E004', 'James O\'Brien', 'Sales', 'Account Executive', 'Junior', 'Chicago', '(312) 555-0104'],
        ['E005', 'Yuki Tanaka', 'Design', 'UX Designer', 'Mid', 'Seattle', '(206) 555-0105'],
        ['E006', 'Olivia Martinez', 'HR', 'Recruiter', 'Junior', 'Denver', '(303) 555-0106'],
        ['E007', 'David Kim', 'Engineering', 'Backend Engineer', 'Mid', 'San Francisco', '(415) 555-0107'],
        ['E008', 'Emma Wilson', 'Finance', 'Financial Analyst', 'Senior', 'New York', '(212) 555-0108'],
        ['E009', 'Raj Gupta', 'Engineering', 'QA Engineer', 'Mid', 'Austin', '(512) 555-0109'],
        ['E010', 'Lauren Foster', 'Marketing', 'Content Strategist', 'Junior', 'Chicago', '(312) 555-0110'],
        ['E011', 'Carlos Rivera', 'Sales', 'Sales Director', 'Lead', 'Miami', '(305) 555-0111'],
        ['E012', 'Hannah Lee', 'Design', 'Visual Designer', 'Junior', 'Seattle', '(206) 555-0112'],
        ['E013', 'Michael Brown', 'Engineering', 'Frontend Engineer', 'Senior', 'San Francisco', '(415) 555-0113'],
        ['E014', 'Sophia Zhang', 'Finance', 'Controller', 'Lead', 'New York', '(212) 555-0114'],
        ['E015', 'Aiden Murphy', 'HR', 'HR Manager', 'Senior', 'Denver', '(303) 555-0115'],
        ['E016', 'Isabella Torres', 'Engineering', 'Data Engineer', 'Mid', 'Austin', '(512) 555-0116'],
        ['E017', 'Nathan Clark', 'Marketing', 'SEO Specialist', 'Junior', 'Chicago', '(312) 555-0117'],
        ['E018', 'Mia Robinson', 'Sales', 'Business Development', 'Mid', 'Miami', '(305) 555-0118'],
        ['E019', 'Ethan Wright', 'Design', 'Product Designer', 'Senior', 'Seattle', '(206) 555-0119'],
        ['E020', 'Ava Hernandez', 'Engineering', 'ML Engineer', 'Senior', 'San Francisco', '(415) 555-0120'],
        ['E021', 'Lucas Scott', 'Finance', 'Budget Analyst', 'Junior', 'New York', '(212) 555-0121'],
        ['E022', 'Chloe Adams', 'HR', 'Training Coordinator', 'Mid', 'Denver', '(303) 555-0122'],
        ['E023', 'Benjamin Park', 'Engineering', 'Site Reliability Eng', 'Mid', 'Austin', '(512) 555-0123'],
        ['E024', 'Grace Nguyen', 'Marketing', 'Brand Manager', 'Senior', 'Chicago', '(312) 555-0124'],
        ['E025', 'Alexander Hall', 'Sales', 'Regional Manager', 'Lead', 'Miami', '(305) 555-0125'],
        ['E026', 'Zoe Campbell', 'Design', 'UI Developer', 'Mid', 'Seattle', '(206) 555-0126'],
        ['E027', 'Daniel Evans', 'Engineering', 'Security Engineer', 'Senior', 'San Francisco', '(415) 555-0127'],
        ['E028', 'Lily Morgan', 'Finance', 'Accounts Payable', 'Junior', 'New York', '(212) 555-0128'],
        ['E029', 'Owen Phillips', 'HR', 'Benefits Specialist', 'Mid', 'Denver', '(303) 555-0129'],
        ['E030', 'Ella Cooper', 'Engineering', 'Platform Engineer', 'Mid', 'Austin', '(512) 555-0130'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # No conditional formatting in initial state
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
