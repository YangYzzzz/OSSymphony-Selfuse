"""
Reward Script: Create named range 'ProductList' for A2:A30 on the Inventory sheet
Task ID: calc_nrv_020
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Named range 'ProductList' exists
  Component 2 (0.3): Named range refers to Inventory!$A$2:$A$30
  Component 3 (0.3): Named range is workbook-scoped and cell data intact
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_020'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Inventory' sheet must exist
    if 'Inventory' not in wb.sheetnames:
        print("FAIL: 'Inventory' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Named range 'ProductList' exists (0.4 points)
    try:
        defined_name = wb.defined_names.get('ProductList')
        if defined_name is not None:
            print(f"PASS: Component 1 — Named range 'ProductList' exists (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — Named range 'ProductList' not found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Named range refers to correct range Inventory!$A$2:$A$30 (0.3 points)
    try:
        defined_name = wb.defined_names.get('ProductList')
        if defined_name is not None:
            ref = defined_name.attr_text
            # Normalize: strip quotes around sheet name if present, e.g. 'Inventory'!$A$2:$A$30
            normalized_ref = ref.replace("'", "")
            expected = "Inventory!$A$2:$A$30"
            if normalized_ref == expected:
                print(f"PASS: Component 2 — Range is '{ref}', matches expected (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — Range is '{ref}', expected '{expected}'")
        else:
            print(f"FAIL: Component 2 — Named range 'ProductList' not found, cannot check range")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Named range is workbook-scoped (localSheetId is None) AND
    # cell data in A2:A30 is intact (0.3 points)
    try:
        defined_name = wb.defined_names.get('ProductList')
        if defined_name is not None:
            # Check workbook scope
            is_workbook_scoped = (defined_name.localSheetId is None)
            if not is_workbook_scoped:
                print(f"FAIL: Component 3 — Named range is sheet-scoped (localSheetId={defined_name.localSheetId}), expected workbook-scoped")
            else:
                # Check data integrity: A2:A30 should have 29 non-empty product names
                ws = wb['Inventory']
                non_empty_count = 0
                for r in range(2, 31):
                    val = ws.cell(row=r, column=1).value
                    if val is not None and str(val).strip() != '':
                        non_empty_count += 1
                if non_empty_count == 29:
                    print(f"PASS: Component 3 — Workbook-scoped, data intact ({non_empty_count} products) (0.3 pts)")
                    total_score += 0.3
                else:
                    print(f"FAIL: Component 3 — Data integrity issue: found {non_empty_count}/29 products in A2:A30")
        else:
            print(f"FAIL: Component 3 — Named range 'ProductList' not found, cannot check scope")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
