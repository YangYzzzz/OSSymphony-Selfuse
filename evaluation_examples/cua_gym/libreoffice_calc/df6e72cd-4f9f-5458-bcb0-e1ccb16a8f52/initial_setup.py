"""
Initial Setup: Create Inventory sheet with 29 product names for named range task
Task ID: calc_nrv_020
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_020'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Inventory sheet ---
    ws = wb.active
    ws.title = 'Inventory'

    # Header
    ws.cell(row=1, column=1, value='Product Name')

    # 29 product names sorted alphabetically (A2:A30)
    products = [
        'Adapter Cable USB-C',
        'Bluetooth Speaker Mini',
        'Ceramic Coffee Mug',
        'Desktop Organizer Tray',
        'Ergonomic Mouse Pad',
        'Fitness Tracker Band',
        'Glass Water Bottle',
        'HDMI Splitter 4-Port',
        'Insulated Lunch Bag',
        'Journal Notebook A5',
        'Keyboard Wrist Rest',
        'LED Desk Lamp',
        'Magnetic Phone Mount',
        'Noise-Canceling Earbuds',
        'Office Chair Cushion',
        'Portable Charger 10000mAh',
        'Quick-Dry Towel Set',
        'Reusable Shopping Bag',
        'Stainless Steel Tumbler',
        'Tablet Stand Adjustable',
        'USB Flash Drive 64GB',
        'Vacuum Insulated Flask',
        'Wireless Charging Pad',
        'XLR Microphone Cable',
        'Yoga Mat Premium',
        'Zinc Alloy Carabiner',
        'Acrylic Desk Shelf',
        'Bamboo Cutting Board',
        'Cotton Canvas Tote',
    ]

    # Sort alphabetically to match context requirement
    products.sort()

    for i, product in enumerate(products, 2):
        ws.cell(row=i, column=1, value=product)

    # Set column width for readability
    ws.column_dimensions['A'].width = 30

    # No named ranges - that is the task for the agent to complete
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
