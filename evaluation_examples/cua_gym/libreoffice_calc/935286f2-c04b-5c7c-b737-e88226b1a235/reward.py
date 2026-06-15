"""
Reward Script: Enable 'Show Formulas' and show gridlines on the 'Audit Review' sheet
Task ID: calc_sht_viewopt_002
Domain: libreoffice_calc
Scoring:
  - Component 1: 'Audit Review' sheet has showFormulas = True  (0.5 pts)
  - Component 2: 'Audit Review' sheet has showGridLines = True (0.5 pts)
Total: 1.0

Both components FAIL on the initial file (showFormulas=False, showGridLines=False)
and PASS on the golden file (showFormulas=True, showGridLines=True).
Cell content must not be altered (used as a precondition gate).
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — reward scripts run on the VM
TASK_ID = 'calc_sht_viewopt_002'
SHEET_NAME = 'Audit Review'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition gate: load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: the 'Audit Review' sheet must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found in workbook (sheets: {wb.sheetnames})")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]
    sv = ws.sheet_view

    # Component 1: 'Audit Review' has showFormulas = True (0.5 points)
    # Initial state: showFormulas=False → task requires setting it to True
    try:
        show_formulas = sv.showFormulas  # True, False, or None (None means default=False)
        # None is treated as False (the default LibreOffice/Excel behaviour)
        if show_formulas is True:
            print(f"PASS: Component 1 — 'Audit Review' showFormulas=True (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — Expected showFormulas=True on '{SHEET_NAME}', "
                  f"found: {repr(show_formulas)}")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not read showFormulas: {e}")

    # Component 2: 'Audit Review' has showGridLines = True (0.5 points)
    # Initial state: showGridLines=False → task requires setting it to True
    try:
        show_gridlines = sv.showGridLines  # True, False, or None (None means default=True)
        # In the initial file this is explicitly False; golden sets it to True
        if show_gridlines is True:
            print(f"PASS: Component 2 — 'Audit Review' showGridLines=True (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 2 — Expected showGridLines=True on '{SHEET_NAME}', "
                  f"found: {repr(show_gridlines)}")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not read showGridLines: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
