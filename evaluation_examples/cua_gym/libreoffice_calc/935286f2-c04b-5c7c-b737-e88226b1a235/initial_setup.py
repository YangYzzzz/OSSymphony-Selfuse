"""
Initial Setup: Enable Show Formulas and show gridlines on Audit Review sheet
Task ID: calc_sht_viewopt_002
Domain: libreoffice_calc

Creates a workbook with three sheets:
- 'Audit Review': complex calculation sheet with formulas,
  currently showing VALUES (showFormulas=False) and has HIDDEN gridlines (showGridLines=False)
- 'Source Data': raw data sheet, gridlines visible (default)
- 'Results': summary sheet, gridlines visible (default)
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

WORKDIR = '/home/user'
TASK_ID = 'calc_sht_viewopt_002'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # -------------------------------------------------------------------------
    # Sheet 1: Source Data
    # -------------------------------------------------------------------------
    ws_src = wb.active
    ws_src.title = 'Source Data'

    src_headers = ['Month', 'Region', 'Product', 'Units Sold', 'Unit Price', 'Discount %', 'Returns']
    for col, h in enumerate(src_headers, 1):
        cell = ws_src.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)

    src_data = [
        ['Jan', 'North', 'Laptop Pro 15',  320,  1299.99, 5.0,  12],
        ['Jan', 'South', 'Laptop Pro 15',  285,  1299.99, 7.5,  9],
        ['Jan', 'East',  'Desktop Ultra',  410,   899.50, 3.0,  15],
        ['Jan', 'West',  'Desktop Ultra',  375,   899.50, 4.0,  11],
        ['Feb', 'North', 'Laptop Pro 15',  298,  1299.99, 5.0,  8],
        ['Feb', 'South', 'Laptop Pro 15',  310,  1299.99, 6.0,  14],
        ['Feb', 'East',  'Desktop Ultra',  389,   899.50, 3.5,  17],
        ['Feb', 'West',  'Desktop Ultra',  402,   899.50, 4.5,  10],
        ['Mar', 'North', 'Laptop Pro 15',  341,  1299.99, 5.0,  7],
        ['Mar', 'South', 'Laptop Pro 15',  327,  1299.99, 7.0,  13],
        ['Mar', 'East',  'Desktop Ultra',  425,   899.50, 3.0,  19],
        ['Mar', 'West',  'Desktop Ultra',  398,   899.50, 4.0,  8],
        ['Apr', 'North', 'Laptop Pro 15',  355,  1299.99, 5.5,  11],
        ['Apr', 'South', 'Laptop Pro 15',  312,  1299.99, 6.5,  9],
        ['Apr', 'East',  'Desktop Ultra',  440,   899.50, 2.5,  14],
        ['Apr', 'West',  'Desktop Ultra',  417,   899.50, 3.5,  12],
        ['May', 'North', 'Laptop Pro 15',  368,  1299.99, 5.0,  10],
        ['May', 'South', 'Laptop Pro 15',  295,  1299.99, 8.0,  16],
        ['May', 'East',  'Desktop Ultra',  456,   899.50, 3.0,  18],
        ['May', 'West',  'Desktop Ultra',  431,   899.50, 4.0,  9],
        ['Jun', 'North', 'Laptop Pro 15',  380,  1299.99, 5.0,  8],
        ['Jun', 'South', 'Laptop Pro 15',  302,  1299.99, 7.0,  11],
        ['Jun', 'East',  'Desktop Ultra',  465,   899.50, 3.0,  15],
        ['Jun', 'West',  'Desktop Ultra',  448,   899.50, 4.5,  13],
        ['Jul', 'North', 'Laptop Pro 15',  395,  1299.99, 5.0,  12],
        ['Jul', 'South', 'Laptop Pro 15',  318,  1299.99, 6.5,  10],
        ['Jul', 'East',  'Desktop Ultra',  478,   899.50, 3.5,  17],
        ['Jul', 'West',  'Desktop Ultra',  462,   899.50, 4.0,  14],
        ['Aug', 'North', 'Laptop Pro 15',  402,  1299.99, 5.0,  9],
        ['Aug', 'South', 'Laptop Pro 15',  335,  1299.99, 7.0,  15],
        ['Aug', 'East',  'Desktop Ultra',  490,   899.50, 2.5,  16],
        ['Aug', 'West',  'Desktop Ultra',  475,   899.50, 3.5,  11],
        ['Sep', 'North', 'Laptop Pro 15',  415,  1299.99, 5.0,  13],
        ['Sep', 'South', 'Laptop Pro 15',  348,  1299.99, 6.0,  8],
        ['Sep', 'East',  'Desktop Ultra',  505,   899.50, 3.0,  19],
        ['Sep', 'West',  'Desktop Ultra',  488,   899.50, 4.0,  10],
        ['Oct', 'North', 'Laptop Pro 15',  428,  1299.99, 5.5,  12],
        ['Oct', 'South', 'Laptop Pro 15',  362,  1299.99, 6.0,  14],
        ['Oct', 'East',  'Desktop Ultra',  518,   899.50, 3.0,  17],
        ['Oct', 'West',  'Desktop Ultra',  501,   899.50, 4.5,  9],
    ]
    for r, row_data in enumerate(src_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_src.cell(row=r, column=c, value=val)

    ws_src.column_dimensions['A'].width = 8
    ws_src.column_dimensions['B'].width = 10
    ws_src.column_dimensions['C'].width = 18
    ws_src.column_dimensions['D'].width = 12
    ws_src.column_dimensions['E'].width = 12
    ws_src.column_dimensions['F'].width = 12
    ws_src.column_dimensions['G'].width = 10
    # Source Data: gridlines visible (default — showGridLines not set / None)

    # -------------------------------------------------------------------------
    # Sheet 2: Results
    # -------------------------------------------------------------------------
    ws_res = wb.create_sheet('Results')

    res_headers = ['Region', 'Product', 'Total Revenue', 'Net Revenue', 'Avg Discount', 'Return Rate %', 'Units Sold']
    for col, h in enumerate(res_headers, 1):
        cell = ws_res.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)

    res_data = [
        ['North', 'Laptop Pro 15',  5318259.36, 5051346.39, 5.10, 3.12, 4096],
        ['South', 'Laptop Pro 15',  4701826.44, 4334863.94, 6.84, 4.10, 3614],
        ['East',  'Desktop Ultra',  3956640.00, 3829213.44, 3.08, 3.94, 4398],
        ['West',  'Desktop Ultra',  3745840.00, 3596978.24, 4.05, 3.24, 4162],
        ['North', 'Desktop Ultra',   784425.50,  752468.48, 3.10, 3.15,  872],
        ['South', 'Desktop Ultra',   698356.25,  659706.44, 4.52, 3.88,  776],
        ['East',  'Laptop Pro 15',  3894103.20, 3778540.10, 5.25, 4.11, 2996],
        ['West',  'Laptop Pro 15',  3572246.40, 3394634.08, 6.15, 3.62, 2748],
    ]
    for r, row_data in enumerate(res_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_res.cell(row=r, column=c, value=val)

    for col_letter, w in zip('ABCDEFG', [12, 18, 16, 16, 14, 14, 12]):
        ws_res.column_dimensions[col_letter].width = w

    # Results: gridlines visible (default)

    # -------------------------------------------------------------------------
    # Sheet 3: Audit Review
    # -------------------------------------------------------------------------
    ws_aud = wb.create_sheet('Audit Review')

    # Column headers (15 columns)
    aud_headers = [
        'Period', 'Region', 'Product', 'Raw Units', 'Unit Price',
        'Gross Revenue', 'Discount %', 'Discount Amt', 'Net Revenue',
        'Returns', 'Return Rate', 'Net Units', 'COGS', 'Gross Profit', 'GP Margin'
    ]
    for col, h in enumerate(aud_headers, 1):
        cell = ws_aud.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)

    # Populate 40 rows with realistic audit data and formulas
    periods = ['Q1-Jan', 'Q1-Feb', 'Q1-Mar', 'Q2-Apr', 'Q2-May', 'Q2-Jun',
               'Q3-Jul', 'Q3-Aug', 'Q3-Sep', 'Q4-Oct']
    regions = ['North', 'South', 'East', 'West']
    products = [('Laptop Pro 15', 1299.99, 0.62), ('Desktop Ultra', 899.50, 0.55)]

    row = 2
    import random
    random.seed(42)
    for period in periods:
        for prod_name, price, cogs_ratio in products:
            for region in regions:
                if row > 41:
                    break
                units = random.randint(280, 520)
                disc = round(random.uniform(2.5, 9.0), 1)

                ws_aud.cell(row=row, column=1, value=period)
                ws_aud.cell(row=row, column=2, value=region)
                ws_aud.cell(row=row, column=3, value=prod_name)
                ws_aud.cell(row=row, column=4, value=units)
                ws_aud.cell(row=row, column=5, value=price)
                # Gross Revenue = Units * Price
                ws_aud.cell(row=row, column=6,  value=f'=D{row}*E{row}')
                # Discount %
                ws_aud.cell(row=row, column=7,  value=disc)
                # Discount Amount = Gross Revenue * Discount% / 100
                ws_aud.cell(row=row, column=8,  value=f'=F{row}*G{row}/100')
                # Net Revenue = Gross Revenue - Discount Amount
                ws_aud.cell(row=row, column=9,  value=f'=F{row}-H{row}')
                # Returns (random count)
                returns = random.randint(5, 25)
                ws_aud.cell(row=row, column=10, value=returns)
                # Return Rate = Returns / Units
                ws_aud.cell(row=row, column=11, value=f'=J{row}/D{row}')
                # Net Units = Units - Returns
                ws_aud.cell(row=row, column=12, value=f'=D{row}-J{row}')
                # COGS = Net Units * Unit Price * COGS ratio (hard-coded as constant per product)
                cogs_per_unit = round(price * cogs_ratio, 2)
                ws_aud.cell(row=row, column=13, value=f'=L{row}*{cogs_per_unit}')
                # Gross Profit = Net Revenue - COGS
                ws_aud.cell(row=row, column=14, value=f'=I{row}-M{row}')
                # GP Margin = Gross Profit / Net Revenue
                ws_aud.cell(row=row, column=15, value=f'=IF(I{row}>0,N{row}/I{row},0)')
                row += 1

    # Totals row
    last_data_row = row - 1
    ws_aud.cell(row=row, column=1, value='TOTAL')
    ws_aud.cell(row=row, column=1).font = Font(bold=True)
    for col_idx in [4, 6, 8, 9, 10, 12, 13, 14]:
        col_letter = get_column_letter(col_idx)
        ws_aud.cell(row=row, column=col_idx,
                    value=f'=SUM({col_letter}2:{col_letter}{last_data_row})')
        ws_aud.cell(row=row, column=col_idx).font = Font(bold=True)

    # Column widths
    for col_idx, w in enumerate([10, 8, 18, 10, 10, 14, 10, 13, 13, 8, 11, 10, 12, 13, 10], 1):
        ws_aud.column_dimensions[get_column_letter(col_idx)].width = w

    # CRITICAL: Audit Review sheet view settings
    # - showFormulas = False  (currently shows values, NOT formulas)
    # - showGridLines = False (gridlines are currently HIDDEN)
    ws_aud.sheet_view.showFormulas = False
    ws_aud.sheet_view.showGridLines = False

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheets: {wb.sheetnames}')
    print(f'  Audit Review: showFormulas=False, showGridLines=False')
    print(f'  Source Data: showFormulas=None (default), showGridLines=None (visible)')
    print(f'  Results: showFormulas=None (default), showGridLines=None (visible)')


create_initial()
