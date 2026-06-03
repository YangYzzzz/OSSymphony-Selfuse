"""
Reward Script: Product Sales Invoice with Itemized Discounts
Task ID: calc_grs_015
Domain: libreoffice_calc
Scoring:
  C1 (0.15) — Sheet named 'Invoice' + merged title 'TAX INVOICE' in A1:G1, bold, centered
  C2 (0.15) — Seller/buyer header sections with merged cells and business details
  C3 (0.15) — Items table header row with 7 columns and fill color
  C4 (0.15) — At least 8 product data rows with codes, names, qty, prices, discounts
  C5 (0.15) — Net Price and Line Total formulas in data rows
  C6 (0.15) — Totals section with Subtotal SUM, Freight, Taxable, GST 10%, Invoice Total formulas
  C7 (0.10) — Footer row with payment instructions in merged cells
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_015'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # ---------------------------------------------------------------
    # Component 1: Sheet named 'Invoice' + merged title 'TAX INVOICE'
    #              in A1 spanning multiple columns, bold, large, centered (0.15 pts)
    # ---------------------------------------------------------------
    try:
        # Must have a sheet (any name accepted, but check for 'Invoice' or similar)
        ws = None
        for name in wb.sheetnames:
            ws_candidate = wb[name]
            a1_val = ws_candidate['A1'].value
            if a1_val and 'TAX INVOICE' in str(a1_val).upper():
                ws = ws_candidate
                break

        if ws is None:
            # Fall back to first sheet
            ws = wb.worksheets[0]
            a1_val = ws['A1'].value

        passed = True
        details = []

        # Check A1 contains 'TAX INVOICE'
        a1_val = ws['A1'].value
        if a1_val and 'TAX INVOICE' in str(a1_val).upper():
            details.append(f"title='{a1_val}'")
        else:
            passed = False
            print(f"FAIL: Component 1 — A1 does not contain 'TAX INVOICE', found: {repr(a1_val)}")

        # Check A1 is merged across multiple columns
        a1_merged = False
        for rng in ws.merged_cells.ranges:
            if rng.min_row == 1 and rng.min_col == 1 and rng.max_col >= 4:
                a1_merged = True
                break
        if not a1_merged:
            passed = False
            print(f"FAIL: Component 1 — A1 is not merged across at least 4 columns")

        # Check bold
        if ws['A1'].font.bold:
            details.append("bold=True")
        else:
            passed = False
            print(f"FAIL: Component 1 — A1 is not bold")

        # Check font size >= 14
        if ws['A1'].font.size and ws['A1'].font.size >= 14:
            details.append(f"size={ws['A1'].font.size}")
        else:
            passed = False
            print(f"FAIL: Component 1 — A1 font size too small: {ws['A1'].font.size}")

        # Check centered alignment
        if ws['A1'].alignment.horizontal == 'center':
            details.append("centered")
        else:
            passed = False
            print(f"FAIL: Component 1 — A1 not centered: {ws['A1'].alignment.horizontal}")

        if passed:
            print(f"PASS: Component 1 — Invoice title correct ({', '.join(details)}) (0.15 pts)")
            total_score += 0.15
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Use the worksheet we found for remaining checks
    if ws is None:
        ws = wb.worksheets[0]

    # ---------------------------------------------------------------
    # Component 2: Seller/buyer header sections with merged cells
    #              and business details (0.15 pts)
    # ---------------------------------------------------------------
    try:
        passed = True
        details = []

        # Check for seller info (left side, rows 3-7ish, columns A-C merged)
        seller_found = False
        buyer_found = False

        # Look for merged ranges in the header area (rows 2-9)
        left_merges = 0
        right_merges = 0
        for rng in ws.merged_cells.ranges:
            if 2 <= rng.min_row <= 9:
                if rng.min_col <= 3 and rng.max_col >= 2:
                    left_merges += 1
                if rng.min_col >= 4 and rng.max_col >= 5:
                    right_merges += 1

        # Check that there are merged cell blocks on left and right
        if left_merges >= 2:
            details.append(f"left_merges={left_merges}")
        else:
            passed = False
            print(f"FAIL: Component 2 — Not enough left-side merged header cells: {left_merges}")

        if right_merges >= 2:
            details.append(f"right_merges={right_merges}")
        else:
            passed = False
            print(f"FAIL: Component 2 — Not enough right-side merged header cells: {right_merges}")

        # Check for seller/buyer text content in rows 3-9
        for row in range(2, 10):
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    val_upper = cell.value.upper()
                    if any(kw in val_upper for kw in ['FROM:', 'SELLER', 'ABN', 'WHOLESALE', 'DISTRIBUTOR']):
                        seller_found = True
                    if any(kw in val_upper for kw in ['TO:', 'BUYER', 'RETAIL', 'CUSTOMER']):
                        buyer_found = True

        if seller_found:
            details.append("seller_info")
        else:
            passed = False
            print(f"FAIL: Component 2 — No seller details found in header area")

        if buyer_found:
            details.append("buyer_info")
        else:
            passed = False
            print(f"FAIL: Component 2 — No buyer details found in header area")

        if passed:
            print(f"PASS: Component 2 — Header sections correct ({', '.join(details)}) (0.15 pts)")
            total_score += 0.15
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---------------------------------------------------------------
    # Component 3: Items table header row with 7 columns and fill color (0.15 pts)
    # ---------------------------------------------------------------
    try:
        # Find the header row — look for a row containing keywords like
        # 'Item Code', 'Product Name', 'Quantity', etc.
        header_row = None
        expected_headers = ['item', 'product', 'quantity', 'price', 'discount', 'net', 'total']

        for row_num in range(8, 20):
            row_values = []
            for col in range(1, 10):
                val = ws.cell(row=row_num, column=col).value
                if val:
                    row_values.append(str(val).lower())
            if len(row_values) >= 5:
                matches = sum(1 for h in expected_headers if any(h in rv for rv in row_values))
                if matches >= 4:
                    header_row = row_num
                    break

        if header_row is None:
            print(f"FAIL: Component 3 — No items table header row found")
        else:
            passed = True
            details = [f"header_row={header_row}"]

            # Count header columns with values
            header_cols = 0
            for col in range(1, 10):
                if ws.cell(row=header_row, column=col).value:
                    header_cols += 1

            if header_cols >= 6:
                details.append(f"cols={header_cols}")
            else:
                passed = False
                print(f"FAIL: Component 3 — Only {header_cols} header columns, expected >= 6")

            # Check fill color on header row
            has_fill = False
            for col in range(1, header_cols + 1):
                cell = ws.cell(row=header_row, column=col)
                if not isinstance(cell, MergedCell):
                    try:
                        fill_rgb = cell.fill.fgColor.rgb
                        if fill_rgb and fill_rgb != '00000000':
                            has_fill = True
                            break
                    except:
                        pass

            if has_fill:
                details.append("has_fill_color")
            else:
                passed = False
                print(f"FAIL: Component 3 — Header row has no fill color")

            if passed:
                print(f"PASS: Component 3 — Header row correct ({', '.join(details)}) (0.15 pts)")
                total_score += 0.15
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ---------------------------------------------------------------
    # Component 4: At least 8 product data rows with item codes,
    #              names, quantities, prices, discounts (0.15 pts)
    # ---------------------------------------------------------------
    try:
        if header_row is None:
            print(f"FAIL: Component 4 — Cannot check data rows without header row")
        else:
            data_rows = 0
            first_data_row = header_row + 1

            for row_num in range(first_data_row, first_data_row + 20):
                # A data row should have at least 5 non-empty cells
                vals = []
                for col in range(1, 10):
                    v = ws.cell(row=row_num, column=col).value
                    if v is not None:
                        vals.append(v)
                # A valid product row has item code, name, qty, price, discount
                if len(vals) >= 5:
                    # Check that there is a numeric quantity and price
                    has_number = False
                    for v in vals:
                        if isinstance(v, (int, float)):
                            has_number = True
                            break
                    if has_number:
                        data_rows += 1
                else:
                    break  # end of data rows

            if data_rows >= 8:
                print(f"PASS: Component 4 — {data_rows} product data rows found (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 4 — Only {data_rows} product data rows, expected >= 8")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ---------------------------------------------------------------
    # Component 5: Net Price and Line Total formulas in data rows (0.15 pts)
    # ---------------------------------------------------------------
    try:
        if header_row is None:
            print(f"FAIL: Component 5 — Cannot check formulas without header row")
        else:
            first_data_row = header_row + 1
            net_price_formulas = 0
            line_total_formulas = 0

            # Find which columns are Net Price and Line Total
            # They should be the last two columns with headers
            net_price_col = None
            line_total_col = None
            for col in range(1, 10):
                hdr = ws.cell(row=header_row, column=col).value
                if hdr:
                    hdr_lower = str(hdr).lower()
                    if 'net' in hdr_lower and 'price' in hdr_lower:
                        net_price_col = col
                    elif 'total' in hdr_lower and 'line' in hdr_lower:
                        line_total_col = col
                    elif 'total' in hdr_lower and net_price_col is not None:
                        # Fallback: if we already found net price, next 'total' is line total
                        if line_total_col is None:
                            line_total_col = col

            # If not found by name, use last two data columns
            if net_price_col is None:
                net_price_col = 6  # F
            if line_total_col is None:
                line_total_col = 7  # G

            for row_num in range(first_data_row, first_data_row + 12):
                np_val = ws.cell(row=row_num, column=net_price_col).value
                lt_val = ws.cell(row=row_num, column=line_total_col).value

                if np_val and isinstance(np_val, str) and np_val.startswith('='):
                    net_price_formulas += 1
                if lt_val and isinstance(lt_val, str) and lt_val.startswith('='):
                    line_total_formulas += 1

            passed = True
            details = []

            if net_price_formulas >= 8:
                details.append(f"net_price_formulas={net_price_formulas}")
            else:
                passed = False
                print(f"FAIL: Component 5 — Only {net_price_formulas} Net Price formulas, expected >= 8")

            if line_total_formulas >= 8:
                details.append(f"line_total_formulas={line_total_formulas}")
            else:
                passed = False
                print(f"FAIL: Component 5 — Only {line_total_formulas} Line Total formulas, expected >= 8")

            if passed:
                print(f"PASS: Component 5 — Data row formulas correct ({', '.join(details)}) (0.15 pts)")
                total_score += 0.15
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # ---------------------------------------------------------------
    # Component 6: Totals section with Subtotal SUM, Freight,
    #              Taxable Amount, GST 10%, Invoice Total formulas (0.15 pts)
    # ---------------------------------------------------------------
    try:
        # Search for totals section below the data rows
        subtotal_found = False
        freight_found = False
        gst_found = False
        invoice_total_found = False
        taxable_found = False

        # Scan rows below data area
        scan_start = (header_row + 9) if header_row else 18
        for row_num in range(scan_start, scan_start + 15):
            for col in range(1, 10):
                cell = ws.cell(row=row_num, column=col)
                if cell.value and isinstance(cell.value, str):
                    val_lower = cell.value.lower().strip()

                    if 'subtotal' in val_lower or 'sub total' in val_lower:
                        # Check for SUM formula in same row, later column
                        for c2 in range(col + 1, 10):
                            v2 = ws.cell(row=row_num, column=c2).value
                            if v2 and isinstance(v2, str) and '=SUM' in v2.upper():
                                subtotal_found = True
                                break

                    if 'freight' in val_lower or 'shipping' in val_lower:
                        # Check that there's a value (numeric or formula) in same row
                        for c2 in range(col + 1, 10):
                            v2 = ws.cell(row=row_num, column=c2).value
                            if v2 is not None:
                                freight_found = True
                                break

                    if 'taxable' in val_lower:
                        for c2 in range(col + 1, 10):
                            v2 = ws.cell(row=row_num, column=c2).value
                            if v2 and isinstance(v2, str) and v2.startswith('='):
                                taxable_found = True
                                break

                    if 'gst' in val_lower or 'tax' in val_lower and '10' in val_lower:
                        for c2 in range(col + 1, 10):
                            v2 = ws.cell(row=row_num, column=c2).value
                            if v2 and isinstance(v2, str) and '0.1' in v2:
                                gst_found = True
                                break

                    if 'invoice total' in val_lower or ('total' in val_lower and 'invoice' in val_lower):
                        for c2 in range(col + 1, 10):
                            v2 = ws.cell(row=row_num, column=c2).value
                            if v2 and isinstance(v2, str) and v2.startswith('='):
                                invoice_total_found = True
                                break

        # Score: need at least 4 of 5 components for full marks
        found_count = sum([subtotal_found, freight_found, taxable_found, gst_found, invoice_total_found])
        details = []
        if subtotal_found: details.append("subtotal_SUM")
        if freight_found: details.append("freight")
        if taxable_found: details.append("taxable_formula")
        if gst_found: details.append("gst_10%")
        if invoice_total_found: details.append("invoice_total_formula")

        if found_count >= 4:
            print(f"PASS: Component 6 — Totals section found ({', '.join(details)}) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 6 — Only {found_count}/5 totals components: {details}")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # ---------------------------------------------------------------
    # Component 7: Footer row with payment instructions in merged cells (0.10 pts)
    # ---------------------------------------------------------------
    try:
        footer_found = False

        # Scan rows from bottom of sheet (use extended range to handle
        # max_row inflation from prior .cell() access)
        for row_num in range(ws.max_row, max(ws.max_row - 15, 1), -1):
            for col in range(1, 4):
                cell = ws.cell(row=row_num, column=col)
                if cell.value and isinstance(cell.value, str):
                    val_lower = cell.value.lower()
                    if any(kw in val_lower for kw in ['payment', 'remit', 'bank', 'bsb', 'account', 'net 30', 'terms']):
                        # Check if this row has merged cells
                        is_in_merge = False
                        for rng in ws.merged_cells.ranges:
                            if rng.min_row <= row_num <= rng.max_row and rng.min_col <= col <= rng.max_col:
                                if rng.max_col - rng.min_col >= 2:  # merged across at least 3 columns
                                    is_in_merge = True
                                    break
                        if is_in_merge:
                            footer_found = True
                            print(f"PASS: Component 7 — Footer with payment instructions found in row {row_num}, merged (0.10 pts)")
                            break
            if footer_found:
                break

        if not footer_found:
            print(f"FAIL: Component 7 — No footer with payment instructions in merged cells found")
        else:
            total_score += 0.10
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    # ---------------------------------------------------------------
    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint — run against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
