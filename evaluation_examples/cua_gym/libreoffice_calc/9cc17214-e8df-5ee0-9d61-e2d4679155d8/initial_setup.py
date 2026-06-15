"""
Initial Setup: Create a spreadsheet with order data, E1='Order Quantity', E2 empty.
Task ID: calc_nrv_072
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_072'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"

    # --- Headers ---
    headers = ['Item', 'Unit Price', 'Quantity', 'Total', 'Order Quantity']
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Data rows ---
    data = [
        ['Wireless Mouse',        24.99,  120, None],
        ['USB-C Hub',             45.50,   85, None],
        ['Mechanical Keyboard',   89.99,   60, None],
        ['Monitor Stand',         34.75,   95, None],
        ['Webcam HD 1080p',       59.00,   40, None],
        ['Noise-Cancelling Headset', 129.99, 30, None],
        ['Laptop Sleeve 15"',     19.95,  200, None],
        ['Desk Lamp LED',         42.00,   70, None],
        ['Ergonomic Chair Pad',   27.50,  110, None],
        ['Cable Management Kit',  15.99,  150, None],
        ['External SSD 1TB',      79.99,   55, None],
        ['Portable Charger',      35.00,   90, None],
    ]

    num_fmt_currency = '$#,##0.00'
    num_fmt_int = '#,##0'

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])  # Item
        price_cell = ws.cell(row=r, column=2, value=row_data[1])  # Unit Price
        price_cell.number_format = num_fmt_currency
        qty_cell = ws.cell(row=r, column=3, value=row_data[2])  # Quantity
        qty_cell.number_format = num_fmt_int
        # Total = Unit Price * Quantity
        total_cell = ws.cell(row=r, column=4)
        total_cell.value = f'=B{r}*C{r}'
        total_cell.number_format = num_fmt_currency
        # Order Quantity (E column) left empty - this is where validation goes

    # --- Column widths ---
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 18

    # --- Alternating row fill ---
    light_fill = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")
    for r in range(2, len(data) + 2):
        if r % 2 == 0:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = light_fill

    # E2 must be empty (no validation, no value)
    # This is where the agent will add validation

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
