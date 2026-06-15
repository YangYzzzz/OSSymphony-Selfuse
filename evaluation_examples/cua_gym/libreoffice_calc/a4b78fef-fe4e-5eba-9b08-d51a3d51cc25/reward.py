"""
Reward Script: Create grand total row at row 15 with SUM formulas, currency formatting, double bottom border, and bold.
Task ID: calc_gsd_004
Domain: libreoffice_calc
Scoring:
  Component 1 — A15 label "Grand Total"          (0.15)
  Component 2 — B15:F15 SUM formulas (rows 2-14) (0.30)
  Component 3 — B15:F15 currency number format    (0.20)
  Component 4 — A15:F15 double bottom border      (0.20)
  Component 5 — A15:F15 bold font                 (0.15)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_004'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    try:
        ws = wb['Expenses']
    except KeyError:
        print("CRITICAL: Sheet 'Expenses' not found")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: A15 contains "Grand Total" (0.15 points)
    try:
        val = ws['A15'].value
        if val is not None and str(val).strip().lower() == 'grand total':
            print(f"PASS: Component 1 — A15 contains '{val}' (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — expected 'Grand Total' in A15, found: {val!r}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: B15:F15 contain SUM formulas covering rows 2-14 (0.30 points)
    # Each correct SUM formula earns 0.06 points (5 cells * 0.06 = 0.30)
    try:
        col_letters = ['B', 'C', 'D', 'E', 'F']
        sum_count = 0
        for col_letter in col_letters:
            cell_ref = f'{col_letter}15'
            val = ws[cell_ref].value
            if val is not None and isinstance(val, str):
                normalized = val.upper().replace(' ', '')
                # Accept =SUM(X2:X14) where X is the column letter
                expected = f'=SUM({col_letter}2:{col_letter}14)'.upper()
                if normalized == expected:
                    sum_count += 1
                    print(f"PASS: Component 2 — {cell_ref} has correct SUM formula: {val}")
                else:
                    print(f"FAIL: Component 2 — {cell_ref} expected {expected}, found: {val!r}")
            else:
                print(f"FAIL: Component 2 — {cell_ref} has no formula, value: {val!r}")
        formula_score = sum_count * 0.06
        if sum_count > 0:
            total_score += formula_score
        print(f"Component 2 subtotal: {sum_count}/5 SUM formulas correct ({formula_score:.2f} pts)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: B15:F15 formatted as currency $#,##0.00 (0.20 points)
    # Each correctly formatted cell earns 0.04 points (5 cells * 0.04 = 0.20)
    try:
        fmt_count = 0
        for col_letter in col_letters:
            cell_ref = f'{col_letter}15'
            nf = ws[cell_ref].number_format
            # Accept formats containing $ and .00 (currency with 2 decimals)
            if nf is not None and '$' in nf and '0.00' in nf:
                fmt_count += 1
                print(f"PASS: Component 3 — {cell_ref} has currency format: {nf!r}")
            else:
                print(f"FAIL: Component 3 — {cell_ref} expected currency format, found: {nf!r}")
        fmt_score = fmt_count * 0.04
        if fmt_count > 0:
            total_score += fmt_score
        print(f"Component 3 subtotal: {fmt_count}/5 cells with currency format ({fmt_score:.2f} pts)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: A15:F15 have double bottom border (0.20 points)
    # Each cell with double bottom border earns ~0.0333 points (6 cells * 0.0333 = 0.20)
    try:
        border_count = 0
        all_cols = ['A', 'B', 'C', 'D', 'E', 'F']
        for col_letter in all_cols:
            cell_ref = f'{col_letter}15'
            border_style = ws[cell_ref].border.bottom.style
            if border_style == 'double':
                border_count += 1
                print(f"PASS: Component 4 — {cell_ref} has double bottom border")
            else:
                print(f"FAIL: Component 4 — {cell_ref} expected double bottom border, found: {border_style!r}")
        border_score = round(border_count * (0.20 / 6), 4)
        if border_count > 0:
            total_score += border_score
        print(f"Component 4 subtotal: {border_count}/6 cells with double bottom border ({border_score:.2f} pts)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: A15:F15 are bold (0.15 points)
    # Each bold cell earns 0.025 points (6 cells * 0.025 = 0.15)
    try:
        bold_count = 0
        for col_letter in all_cols:
            cell_ref = f'{col_letter}15'
            is_bold = ws[cell_ref].font.bold
            if is_bold:
                bold_count += 1
                print(f"PASS: Component 5 — {cell_ref} is bold")
            else:
                print(f"FAIL: Component 5 — {cell_ref} expected bold, found: {is_bold!r}")
        bold_score = round(bold_count * (0.15 / 6), 4)
        if bold_count > 0:
            total_score += bold_score
        print(f"Component 5 subtotal: {bold_count}/6 bold cells ({bold_score:.2f} pts)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score:.4f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
