"""
Initial Setup: Expense summary spreadsheet with 13 departments, raw numbers, no totals.
Task ID: calc_gsd_004
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_004'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Expenses'

    # Headers in row 1
    headers = ['Department', 'Q1', 'Q2', 'Q3', 'Q4', 'Annual Total']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # 13 departments with realistic expense data (rows 2-14)
    departments = [
        ['Marketing',        125400, 138200, 142800, 155600, 562000],
        ['Engineering',      245000, 251300, 268700, 274500, 1039500],
        ['HR',                78500,  82100,  79300,  85200,  325100],
        ['Sales',            198700, 205400, 212600, 221800,  838500],
        ['Finance',           95200,  97800, 101400, 103600,  398000],
        ['Legal',             67300,  69800,  72400,  74100,  283600],
        ['Operations',       156800, 162300, 168900, 175400,  663400],
        ['IT',               134600, 139200, 145800, 150300,  569900],
        ['R&D',              312500, 325800, 338600, 352100, 1329000],
        ['Customer Support',  89400,  92700,  96100,  99800,  378000],
        ['Logistics',        112300, 116500, 121400, 126700,  476900],
        ['Facilities',        54800,  56200,  58900,  61300,  231200],
        ['Executive',        185000, 188500, 192700, 197300,  763500],
    ]

    for r, row_data in enumerate(departments, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Row 15 is intentionally left empty - agent must add grand total here

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 20
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col_letter].width = 15

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
