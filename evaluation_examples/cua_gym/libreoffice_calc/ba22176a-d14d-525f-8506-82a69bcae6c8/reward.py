"""
Reward Script: Calculate future value of a lump sum investment
Task ID: calc_fmb_fv_lump_sum_059
Domain: libreoffice_calc

Scoring Rubric:
  Component 1 (0.4): B6 contains an FV formula (not empty, starts with =FV)
  Component 2 (0.4): The FV formula has correct cell references for monthly compounding
                     (rate=B3/B5, nper=B4*B5, pmt=0, pv=-B2)
  Component 3 (0.2): Input cells B2-B5 are unchanged from expected values

Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — reward scripts run on the VM
TASK_ID = 'calc_fmb_fv_lump_sum_059'
SHEET_NAME = 'Lump Sum Calculator'


def normalize_formula(formula):
    """Normalize formula for comparison: uppercase, remove spaces."""
    if not isinstance(formula, str):
        return ''
    return formula.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: sheet must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Component 1: B6 contains an FV formula (0.4 points)
    # This FAILS on initial (B6=None) and PASSES on golden (B6='=FV(...)')
    try:
        b6_value = ws['B6'].value
        if b6_value is not None and isinstance(b6_value, str) and normalize_formula(b6_value).startswith('=FV('):
            print(f"PASS: Component 1 — B6 contains an FV formula: {repr(b6_value)} (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — B6 expected an FV formula, found: {repr(b6_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not check B6: {e}")

    # Component 2: The FV formula uses correct cell references for monthly compounding (0.4 points)
    # Expected: =FV(B3/B5, B4*B5, 0, -B2)  (case-insensitive, whitespace-insensitive)
    # This FAILS on initial (B6=None) and PASSES on golden (correct formula)
    try:
        b6_value = ws['B6'].value
        if b6_value is not None and isinstance(b6_value, str):
            norm = normalize_formula(b6_value)
            # Check for exact formula structure
            # Acceptable forms: =FV(B3/B5,B4*B5,0,-B2) or =FV(B3/B5,B4*B5,0,-B2,0)
            exact_match = norm in ('=FV(B3/B5,B4*B5,0,-B2)', '=FV(B3/B5,B4*B5,0,-B2,0)')
            if exact_match:
                print(f"PASS: Component 2 — FV formula has correct structure: rate=B3/B5, nper=B4*B5, pmt=0, pv=-B2 (0.4 pts)")
                total_score += 0.4
            else:
                # Partial check: verify key elements are present
                has_rate = 'B3/B5' in norm
                has_nper = 'B4*B5' in norm
                has_pv = '-B2' in norm
                has_pmt_zero = ',0,' in norm or ',0)' in norm
                if has_rate and has_nper and has_pv and has_pmt_zero:
                    print(f"PASS: Component 2 — FV formula contains correct elements (rate=B3/B5, nper=B4*B5, pmt=0, pv=-B2): {repr(b6_value)} (0.4 pts)")
                    total_score += 0.4
                else:
                    print(f"FAIL: Component 2 — FV formula does not have correct structure.")
                    print(f"  Found: {repr(b6_value)}")
                    print(f"  Expected structure: =FV(B3/B5, B4*B5, 0, -B2)")
                    print(f"  has_rate(B3/B5)={has_rate}, has_nper(B4*B5)={has_nper}, has_pmt_zero={has_pmt_zero}, has_pv(-B2)={has_pv}")
        else:
            print(f"FAIL: Component 2 — B6 is empty or not a formula string, cannot verify structure.")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not check formula structure: {e}")

    # Component 3: Input cells B2-B5 are unchanged (0.2 points)
    # Expected: B2=10000, B3=0.06, B4=20, B5=12
    # These values should be the same in initial and golden; this serves as data integrity check.
    # NOTE: If these are the same in both initial and golden, we use as compound check
    # embedded with task context (the data must still be correct for the formula to work correctly).
    # Per rules: only score if this FAILS on initial and PASSES on golden.
    # However, B2-B5 are expected to be the same in both. Let's make this a compound check:
    # "B6 formula exists AND input data is intact" — but since Component 1 already checks B6 exists,
    # we'll verify that the referenced cells have the expected values as a data integrity gate
    # and add it as a sub-check to ensure the golden file hasn't corrupted the inputs.
    #
    # To avoid scoring pre-existing properties, this component only passes if
    # B6 has the formula AND the inputs are correct (compound condition anchored to task change).
    try:
        b6_has_formula = ws['B6'].value is not None and isinstance(ws['B6'].value, str) and normalize_formula(ws['B6'].value).startswith('=FV(')
        b2_ok = ws['B2'].value == 10000
        b3_ok = ws['B3'].value is not None and abs(float(ws['B3'].value) - 0.06) < 1e-9
        b4_ok = ws['B4'].value == 20
        b5_ok = ws['B5'].value == 12

        if b6_has_formula and b2_ok and b3_ok and b4_ok and b5_ok:
            print(f"PASS: Component 3 — FV formula present AND input data intact (B2=10000, B3=0.06, B4=20, B5=12) (0.2 pts)")
            total_score += 0.2
        elif not b6_has_formula:
            print(f"FAIL: Component 3 — B6 has no FV formula (precondition for Component 3 not met)")
        else:
            issues = []
            if not b2_ok:
                issues.append(f"B2={repr(ws['B2'].value)} (expected 10000)")
            if not b3_ok:
                issues.append(f"B3={repr(ws['B3'].value)} (expected 0.06)")
            if not b4_ok:
                issues.append(f"B4={repr(ws['B4'].value)} (expected 20)")
            if not b5_ok:
                issues.append(f"B5={repr(ws['B5'].value)} (expected 12)")
            print(f"FAIL: Component 3 — Input data modified: {', '.join(issues)}")
    except Exception as e:
        print(f"ERROR: Component 3 — Could not check input cells: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
