"""
Initial Setup: Future Value Lump Sum Calculator
Task ID: calc_fmb_fv_lump_sum_059
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_fv_lump_sum_059'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Lump Sum Calculator ---
    ws = wb.active
    ws.title = 'Lump Sum Calculator'

    # Column widths for readability
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18

    # Title row
    ws['A1'] = 'Lump Sum Future Value Calculator'
    ws['A1'].font = Font(bold=True, size=13)
    ws.merge_cells('A1:B1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Input labels and values
    ws['A2'] = 'Initial Investment (PV)'
    ws['B2'] = 10000
    ws['B2'].number_format = '$#,##0.00'

    ws['A3'] = 'Annual Interest Rate'
    ws['B3'] = 0.06
    ws['B3'].number_format = '0.00%'

    ws['A4'] = 'Years'
    ws['B4'] = 20

    ws['A5'] = 'Compounding (per year)'
    ws['B5'] = 12

    # Output row — B6 intentionally left empty (target cell)
    ws['A6'] = 'Future Value'
    # B6 is the target cell — must be empty in initial state

    # Style label column bold
    for row in range(2, 7):
        ws.cell(row=row, column=1).font = Font(bold=True)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
