"""
Initial Setup: Set up validation on B2 for US phone numbers
Task ID: calc_nrv_078
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_078'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Contacts"

    # --- Headers ---
    headers = ['Contact Name', 'Phone']
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="000000")
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border

    # --- Contact Data (B2 is intentionally left EMPTY for the task) ---
    # We put data starting from row 3 so B2 remains the target cell
    contacts = [
        # (Name, Phone) - B2 is empty, data starts at row 3
        ['Sarah Chen', '415-555-0142'],
        ['Marcus Johnson', '212-555-0198'],
        ['Elena Rodriguez', '305-555-0267'],
        ['David Kim', '650-555-0334'],
        ['Rachel Thompson', '773-555-0421'],
        ['James O\'Brien', '617-555-0503'],
        ['Priya Patel', '408-555-0687'],
        ['Michael Washington', '310-555-0759'],
        ['Lisa Nakamura', '206-555-0845'],
        ['Carlos Mendez', '512-555-0913'],
        ['Amanda Foster', '202-555-0178'],
        ['Robert Chang', '415-555-0256'],
    ]

    # B2 stays empty -- this is where the agent needs to set up validation
    # Write existing contacts starting from row 3
    for r, (name, phone) in enumerate(contacts, 3):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=phone)

    # Apply light formatting to data rows
    data_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    for r in range(2, len(contacts) + 3):
        for c in range(1, 3):
            cell = ws.cell(row=r, column=c)
            cell.border = data_border
            cell.font = Font(name="Calibri", size=11)

    # Set column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
