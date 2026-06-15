"""
Reward Script: Set up text length validation on B2 for US phone numbers (XXX-XXX-XXXX, 12 chars)
Task ID: calc_nrv_078
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): Data validation exists on B2 with type=textLength
  Component 2 (0.3): Operator is "equal" and formula1 is "12"
  Component 3 (0.2): Input message configured with phone format hint
  Component 4 (0.2): Error message configured
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_078'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Precondition: headers exist (gate, not scored)
    if ws['A1'].value != 'Contact Name' or ws['B1'].value != 'Phone':
        print("WARN: Headers may have been modified, continuing verification anyway")

    # Find data validation that covers B2
    dv_list = ws.data_validations.dataValidation
    target_dv = None
    for dv in dv_list:
        sqref_str = str(dv.sqref)
        # Check if B2 is in the sqref range
        if 'B2' in sqref_str:
            target_dv = dv
            break

    # Component 1: Data validation exists on B2 with type=textLength (0.3 points)
    try:
        if target_dv is not None and target_dv.type == 'textLength':
            print(f"PASS: Component 1 — textLength validation found on B2 (0.3 pts)")
            total_score += 0.3
        elif target_dv is not None:
            print(f"FAIL: Component 1 — validation exists on B2 but type is '{target_dv.type}', expected 'textLength'")
        else:
            print(f"FAIL: Component 1 — no data validation found covering B2")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Operator is "equal" and formula1 is "12" (0.3 points)
    try:
        if target_dv is not None:
            op_ok = target_dv.operator == 'equal'
            f1_ok = str(target_dv.formula1).strip() == '12'
            if op_ok and f1_ok:
                print(f"PASS: Component 2 — operator='equal', formula1='12' (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — operator='{target_dv.operator}' (expected 'equal'), formula1='{target_dv.formula1}' (expected '12')")
        else:
            print(f"FAIL: Component 2 — no data validation found on B2")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Input message configured with phone format hint (0.2 points)
    try:
        if target_dv is not None:
            has_input_msg = (
                target_dv.showInputMessage is True
                and target_dv.prompt is not None
                and len(str(target_dv.prompt).strip()) > 0
            )
            # Check that the prompt mentions the phone format pattern
            prompt_text = str(target_dv.prompt).strip().upper() if target_dv.prompt else ''
            mentions_format = 'XXX' in prompt_text or 'PHONE' in prompt_text or '12' in prompt_text
            if has_input_msg and mentions_format:
                print(f"PASS: Component 3 — input message configured: '{target_dv.prompt}' (0.2 pts)")
                total_score += 0.2
            elif has_input_msg:
                # Input message exists but doesn't mention format — give partial
                print(f"PARTIAL: Component 3 — input message exists but doesn't mention phone format: '{target_dv.prompt}' (0.1 pts)")
                total_score += 0.1
            else:
                print(f"FAIL: Component 3 — showInputMessage={target_dv.showInputMessage}, prompt='{target_dv.prompt}'")
        else:
            print(f"FAIL: Component 3 — no data validation found on B2")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Error message configured (0.2 points)
    try:
        if target_dv is not None:
            has_error_msg = (
                target_dv.showErrorMessage is True
                and target_dv.error is not None
                and len(str(target_dv.error).strip()) > 0
            )
            if has_error_msg:
                print(f"PASS: Component 4 — error message configured: '{target_dv.error}' (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 4 — showErrorMessage={target_dv.showErrorMessage}, error='{target_dv.error}'")
        else:
            print(f"FAIL: Component 4 — no data validation found on B2")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
