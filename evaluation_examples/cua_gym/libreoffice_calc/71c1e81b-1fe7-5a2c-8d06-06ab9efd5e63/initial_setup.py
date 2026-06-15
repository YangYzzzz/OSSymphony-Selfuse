"""
Initial Setup: Create sales data spreadsheet with 1000 transactions across 50 products
Task ID: calc_gcp_057
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time
from datetime import datetime, timedelta

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_057'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

# --- 50 unique products with categories ---
PRODUCTS = [
    ("Wireless Bluetooth Headphones", "Electronics"),
    ("Organic Green Tea", "Beverages"),
    ("Stainless Steel Water Bottle", "Kitchen"),
    ("Yoga Mat Premium", "Fitness"),
    ("LED Desk Lamp", "Office"),
    ("Running Shoes Pro", "Footwear"),
    ("Laptop Stand Adjustable", "Office"),
    ("Protein Powder Vanilla", "Health"),
    ("Ceramic Coffee Mug", "Kitchen"),
    ("USB-C Hub 7-in-1", "Electronics"),
    ("Bamboo Cutting Board", "Kitchen"),
    ("Resistance Bands Set", "Fitness"),
    ("Noise Cancelling Earbuds", "Electronics"),
    ("Herbal Shampoo", "Personal Care"),
    ("Mechanical Keyboard", "Electronics"),
    ("Standing Desk Converter", "Office"),
    ("Cast Iron Skillet", "Kitchen"),
    ("Foam Roller", "Fitness"),
    ("Portable Charger 20000mAh", "Electronics"),
    ("Almond Butter Organic", "Food"),
    ("Hiking Backpack 40L", "Outdoors"),
    ("Wireless Mouse Ergonomic", "Electronics"),
    ("Essential Oil Diffuser", "Home"),
    ("Cotton Bath Towel Set", "Home"),
    ("Smart Watch Fitness", "Electronics"),
    ("Blender High Speed", "Kitchen"),
    ("Vitamin D3 Supplements", "Health"),
    ("Canvas Tote Bag", "Accessories"),
    ("Electric Toothbrush", "Personal Care"),
    ("Desk Organizer Wood", "Office"),
    ("Air Purifier HEPA", "Home"),
    ("Insulated Lunch Box", "Kitchen"),
    ("Cycling Shorts Padded", "Fitness"),
    ("Webcam HD 1080p", "Electronics"),
    ("Natural Deodorant", "Personal Care"),
    ("Instant Coffee Blend", "Beverages"),
    ("Garden Tool Set", "Outdoors"),
    ("Weighted Jump Rope", "Fitness"),
    ("Notebook Leather Bound", "Office"),
    ("Sunscreen SPF 50", "Personal Care"),
    ("Dumbbell Set Adjustable", "Fitness"),
    ("Wireless Charger Pad", "Electronics"),
    ("Coconut Oil Organic", "Food"),
    ("Rain Jacket Waterproof", "Outdoors"),
    ("Monitor Arm Single", "Office"),
    ("Stainless Steel Thermos", "Kitchen"),
    ("Resistance Band Heavy", "Fitness"),
    ("Blue Light Glasses", "Accessories"),
    ("Matcha Powder Premium", "Beverages"),
    ("Hand Cream Moisturizing", "Personal Care"),
]


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)  # Reproducible data

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AllProducts"

    # Headers
    headers = ["TransID", "Date", "ProductName", "Category", "Quantity", "Revenue"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Generate 1000 transactions
    start_date = datetime(2025, 1, 1)
    end_date = datetime(2025, 12, 31)
    date_range = (end_date - start_date).days

    # We want the top product to be approximately $85,000 in total revenue.
    # With 1000 transactions across 50 products, avg ~20 transactions per product.
    # We'll weight product selection to create a realistic distribution.
    # Higher-index products get slightly less weight to create natural variation.

    # Assign weights per product to control transaction frequency
    product_weights = []
    for i in range(50):
        if i < 1:
            product_weights.append(8.0)    # top seller
        elif i < 3:
            product_weights.append(5.0)    # high tier
        elif i < 5:
            product_weights.append(4.0)    # upper-mid tier
        elif i < 10:
            product_weights.append(2.5)    # mid tier
        elif i < 20:
            product_weights.append(1.0)    # lower-mid tier
        elif i < 35:
            product_weights.append(0.5)    # low tier
        else:
            product_weights.append(0.3)    # bottom tier

    # Revenue per transaction ranges (to control totals)
    revenue_ranges = []
    for i in range(50):
        if i < 1:
            revenue_ranges.append((450, 1000))  # premium items
        elif i < 3:
            revenue_ranges.append((300, 700))   # high revenue
        elif i < 5:
            revenue_ranges.append((250, 600))   # upper-mid revenue
        elif i < 10:
            revenue_ranges.append((150, 400))   # mid revenue
        elif i < 20:
            revenue_ranges.append((80, 250))    # lower-mid revenue
        elif i < 35:
            revenue_ranges.append((40, 150))    # low revenue
        else:
            revenue_ranges.append((20, 80))     # bottom revenue

    for trans_id in range(1, 1001):
        # Weighted product selection
        product_idx = random.choices(range(50), weights=product_weights, k=1)[0]
        product_name, category = PRODUCTS[product_idx]

        # Random date
        day_offset = random.randint(0, date_range)
        trans_date = start_date + timedelta(days=day_offset)

        # Quantity and revenue
        quantity = random.randint(1, 10)
        rev_low, rev_high = revenue_ranges[product_idx]
        revenue = round(random.uniform(rev_low, rev_high), 2)

        ws.cell(row=trans_id + 1, column=1, value=trans_id)
        ws.cell(row=trans_id + 1, column=2, value=trans_date.strftime("%Y-%m-%d"))
        ws.cell(row=trans_id + 1, column=3, value=product_name)
        ws.cell(row=trans_id + 1, column=4, value=category)
        ws.cell(row=trans_id + 1, column=5, value=quantity)
        ws.cell(row=trans_id + 1, column=6, value=revenue)

    # Set column widths for readability
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
