"""
Reward Script: Create pivot table from sales data with top 10 products by revenue
Task ID: calc_gcp_057
Domain: libreoffice_calc

Scoring Rubric:
  Component 1 (0.25): PivotTable sheet exists
  Component 2 (0.25): PivotTable has correct structure (ProductName + Sum of Revenue columns, ~50 product rows)
  Component 3 (0.25): Products sorted descending by revenue, top product ~$88k
  Component 4 (0.25): Exactly 10 visible product rows, remaining 40 hidden
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_057'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: PivotTable sheet exists (0.25 points)
    # This is a task-introduced change: initial has only AllProducts.
    try:
        pivot_ws = None
        for name in wb.sheetnames:
            if 'pivot' in name.lower():
                pivot_ws = wb[name]
                break
        if pivot_ws is not None:
            print(f"PASS: Component 1 — PivotTable sheet found: '{pivot_ws.title}' (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — No sheet with 'pivot' in name. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 2: PivotTable has correct structure (0.25 points)
    # Should have ProductName and Revenue-related columns, with ~50 product rows
    try:
        header_a = pivot_ws.cell(row=1, column=1).value
        header_b = pivot_ws.cell(row=1, column=2).value

        has_product_header = header_a is not None and 'product' in str(header_a).lower()
        has_revenue_header = header_b is not None and 'revenue' in str(header_b).lower()

        # Count total data rows (non-empty A column, starting from row 2)
        total_data_rows = 0
        for r in range(2, pivot_ws.max_row + 1):
            val = pivot_ws.cell(row=r, column=1).value
            if val is not None and str(val).strip() != '':
                total_data_rows += 1

        structure_ok = has_product_header and has_revenue_header and (40 <= total_data_rows <= 55)

        if structure_ok:
            print(f"PASS: Component 2 — Headers: '{header_a}', '{header_b}'; {total_data_rows} product rows (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — product_header={has_product_header} ({header_a}), "
                  f"revenue_header={has_revenue_header} ({header_b}), data_rows={total_data_rows}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Products sorted descending by revenue (0.25 points)
    # Top product should have revenue ~$85k-$90k. All visible rows should be in descending order.
    try:
        revenues = []
        for r in range(2, pivot_ws.max_row + 1):
            product = pivot_ws.cell(row=r, column=1).value
            rev = pivot_ws.cell(row=r, column=2).value
            if product is not None and rev is not None:
                try:
                    revenues.append(float(rev))
                except (ValueError, TypeError):
                    pass

        if len(revenues) >= 10:
            # Check if sorted descending (allow small tolerance for ties)
            is_sorted = all(revenues[i] >= revenues[i+1] - 0.01 for i in range(len(revenues) - 1))
            top_revenue = revenues[0]
            # Top product should be approximately $85k-$90k based on context
            top_in_range = 50000 <= top_revenue <= 120000

            if is_sorted and top_in_range:
                print(f"PASS: Component 3 — Sorted descending, top revenue: ${top_revenue:,.2f} (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 3 — sorted={is_sorted}, top_revenue=${top_revenue:,.2f}, top_in_range={top_in_range}")
        else:
            print(f"FAIL: Component 3 — Only {len(revenues)} revenue values found, need at least 10")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Exactly 10 visible product rows, remaining ~40 hidden (0.25 points)
    # This is the core "top 10" filtering requirement.
    try:
        visible_count = 0
        hidden_count = 0
        for r in range(2, pivot_ws.max_row + 1):
            product = pivot_ws.cell(row=r, column=1).value
            if product is not None and str(product).strip() != '':
                if pivot_ws.row_dimensions[r].hidden:
                    hidden_count += 1
                else:
                    visible_count += 1

        if visible_count == 10 and hidden_count >= 35:
            print(f"PASS: Component 4 — {visible_count} visible, {hidden_count} hidden product rows (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 4 — {visible_count} visible (expected 10), {hidden_count} hidden (expected ~40)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
