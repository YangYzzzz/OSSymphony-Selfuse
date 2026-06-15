"""
Reward Script: Fill Address, Website, and Phone for 5 Korean BBQ restaurants in SEOUL_KBBQ.xlsx
Task ID: osworld_multi_apps_restaurant_lookup_009
Domain: libreoffice_calc
Scoring:
  - Component 1: Address column filled for all 5 restaurants (0.4 pts)
  - Component 2: Phone column filled for all 5 restaurants (0.3 pts)
  - Component 3: At least one website entry filled (0.1 pts)
  - Component 4: At least 4 restaurants have both Address AND Phone filled (0.2 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_restaurant_lookup_009'
FILE_PATH = f'{WORKDIR}/SEOUL_KBBQ.xlsx'

# Expected restaurant names in rows 2-6 (column A)
EXPECTED_NAMES = [
    'Maple Tree House Itaewon',
    'Palsaik Samgyeopsal',
    'Changssam',
    'Hongdae Gopchang',
    'Bornga Sinchon',
]

def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Checks that the agent filled in Address, Website, and Phone columns
    in SEOUL_KBBQ.xlsx based on Google Maps data.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get the active/first sheet
    try:
        ws = wb.active
        print(f"Sheet: {ws.title}")
    except Exception as e:
        print(f"CRITICAL: Cannot access worksheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify header row as precondition gate
    try:
        header_name = ws['A1'].value
        header_addr = ws['B1'].value
        header_web  = ws['C1'].value
        header_phone = ws['D1'].value
        expected_headers = ('Name', 'Address', 'Website', 'Phone')
        actual_headers = (header_name, header_addr, header_web, header_phone)
        if actual_headers != expected_headers:
            print(f"WARN: Unexpected headers: {actual_headers}, expected {expected_headers}")
    except Exception as e:
        print(f"WARN: Could not verify headers: {e}")

    # Collect data rows 2-6 (the 5 restaurants)
    rows = []
    try:
        for row_idx in range(2, 7):
            name = ws.cell(row=row_idx, column=1).value
            address = ws.cell(row=row_idx, column=2).value
            website = ws.cell(row=row_idx, column=3).value
            phone   = ws.cell(row=row_idx, column=4).value
            rows.append((name, address, website, phone))
        print(f"Data rows collected: {len(rows)}")
        for i, row in enumerate(rows, 1):
            print(f"  Row {i+1}: {row}")
    except Exception as e:
        print(f"CRITICAL: Could not read data rows: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Address column filled for all 5 restaurants (0.4 points)
    # Initial state has all None in Address column; golden has all 5 filled.
    try:
        addresses_filled = sum(
            1 for (_, addr, _, _) in rows
            if addr is not None and str(addr).strip() != ''
        )
        if addresses_filled == 5:
            print(f"PASS: Component 1 — All 5 addresses filled ({addresses_filled}/5) (0.4 pts)")
            total_score += 0.4
        elif addresses_filled >= 3:
            partial = 0.2
            print(f"PARTIAL: Component 1 — {addresses_filled}/5 addresses filled ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {addresses_filled}/5 addresses filled (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Phone column filled for all 5 restaurants (0.3 points)
    # Initial state has all None in Phone column; golden has all 5 filled.
    try:
        phones_filled = sum(
            1 for (_, _, _, phone) in rows
            if phone is not None and str(phone).strip() != ''
        )
        if phones_filled == 5:
            print(f"PASS: Component 2 — All 5 phones filled ({phones_filled}/5) (0.3 pts)")
            total_score += 0.3
        elif phones_filled >= 3:
            partial = 0.15
            print(f"PARTIAL: Component 2 — {phones_filled}/5 phones filled ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {phones_filled}/5 phones filled (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: At least one website entry filled (0.1 points)
    # Initial state has all None; golden has at least 3 filled.
    try:
        websites_filled = sum(
            1 for (_, _, web, _) in rows
            if web is not None and str(web).strip() != ''
        )
        if websites_filled >= 1:
            print(f"PASS: Component 3 — {websites_filled} website(s) filled (0.1 pts)")
            total_score += 0.1
        else:
            print(f"FAIL: Component 3 — No websites filled (expected at least 1)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: At least 4 restaurants have both Address AND Phone filled (0.2 points)
    # This checks compound completeness — that agent filled both key fields per restaurant.
    # Initial state: 0 restaurants have either; golden: all 5 have Address + Phone.
    try:
        both_filled = sum(
            1 for (_, addr, _, phone) in rows
            if (addr is not None and str(addr).strip() != '')
            and (phone is not None and str(phone).strip() != '')
        )
        if both_filled >= 4:
            print(f"PASS: Component 4 — {both_filled}/5 restaurants have both Address+Phone (0.2 pts)")
            total_score += 0.2
        elif both_filled >= 2:
            partial = 0.1
            print(f"PARTIAL: Component 4 — {both_filled}/5 restaurants have both Address+Phone ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {both_filled}/5 restaurants have both Address+Phone (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


if not os.path.exists(FILE_PATH):
    print(f"File not found: {FILE_PATH}")
    print("REWARD: 0.0")
else:
    verify_task(FILE_PATH)
