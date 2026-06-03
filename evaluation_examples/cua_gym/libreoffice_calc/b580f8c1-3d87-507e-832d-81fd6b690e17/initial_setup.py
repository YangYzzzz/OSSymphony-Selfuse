"""
Initial Setup: Seoul Korean BBQ Restaurant Lookup
Task ID: osworld_multi_apps_restaurant_lookup_009
Domain: libreoffice_calc (multi-app: gedit + LibreOffice Calc)

Creates:
  - /home/user/Desktop/seoul_kbbq.txt  (restaurant name list, opened in gedit)
  - /home/user/SEOUL_KBBQ.xlsx         (spreadsheet with Name column pre-filled, opened in LibreOffice Calc)
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
DESKTOP = '/home/user/Desktop'
TASK_ID = 'osworld_multi_apps_restaurant_lookup_009'
TXT_FILE = f'{DESKTOP}/seoul_kbbq.txt'
XLSX_FILE = f'{WORKDIR}/SEOUL_KBBQ.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch a GUI app on the VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    # Ensure Desktop directory exists
    os.makedirs(DESKTOP, exist_ok=True)

    # --- Create seoul_kbbq.txt on the Desktop ---
    restaurant_names = [
        "Maple Tree House Itaewon",
        "Palsaik Samgyeopsal",
        "Changssam",
        "Hongdae Gopchang",
        "Bornga Sinchon",
    ]
    with open(TXT_FILE, 'w', encoding='utf-8') as f:
        f.write('\n'.join(restaurant_names) + '\n')
    print(f'Text file created: {TXT_FILE}')

    # --- Create SEOUL_KBBQ.xlsx with Name column pre-filled ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # Column headers in row 1
    headers = ['Name', 'Address', 'Website', 'Phone']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Restaurant names pre-filled in column A, rows 2-6
    # Address, Website, Phone columns are left empty (task is to fill these)
    for row_idx, name in enumerate(restaurant_names, 2):
        ws.cell(row=row_idx, column=1, value=name)
        # Columns B (Address), C (Website), D (Phone) intentionally left blank

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 20

    wb.save(XLSX_FILE)
    print(f'Spreadsheet created: {XLSX_FILE}')

    # --- GUI-ready startup ---
    # Open seoul_kbbq.txt in gedit
    launch_gui(f'gedit "{TXT_FILE}"', delay_sec=2.0)

    # Open SEOUL_KBBQ.xlsx in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{XLSX_FILE}"', delay_sec=2.0)

    print('GUI_READY: launched gedit and LibreOffice Calc with DISPLAY=:0')


create_initial()
