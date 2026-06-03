"""
Reward Script: Payroll summary report with formulas, formatting, and conditional formatting
Task ID: calc_gsd_044
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25) - Row 52 totals: 'Payroll Total' in A52 + SUM formulas in D52:K52
  Component 2 (0.15) - Bold formatting on header row 1 and totals row 52
  Component 3 (0.20) - USD currency number format ($#,##0.00) on D2:K52
  Component 4 (0.15) - Conditional formatting on K2:K51 (lessThan 2000, red background)
  Component 5 (0.15) - Thin borders on A1:K52
  Component 6 (0.10) - Freeze panes at A2 (row 1 frozen)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_044'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    try:
        ws = wb['Payroll']
    except KeyError:
        print("CRITICAL: Sheet 'Payroll' not found")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Row 52 totals row (0.25 points)
    # A52 = 'Payroll Total', D52:K52 have SUM formulas
    try:
        a52_val = ws.cell(row=52, column=1).value
        label_ok = (a52_val is not None and
                    str(a52_val).strip().lower() == 'payroll total')

        # Check SUM formulas in columns D(4) through K(11) in row 52
        sum_cols = list(range(4, 12))  # D=4, E=5, ..., K=11
        col_letters = {4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K'}
        formulas_ok = 0
        for c in sum_cols:
            val = ws.cell(row=52, column=c).value
            if val is not None and isinstance(val, str):
                normalized = val.upper().replace(' ', '')
                expected = f'=SUM({col_letters[c]}2:{col_letters[c]}51)'
                if normalized == expected:
                    formulas_ok += 1
                else:
                    print(f"  INFO: R52C{c} formula={val}, expected {expected}")
            else:
                print(f"  INFO: R52C{c} value={val}, expected SUM formula")

        if label_ok and formulas_ok == 8:
            print(f"PASS: Component 1 — Row 52 has 'Payroll Total' and all 8 SUM formulas (0.25 pts)")
            total_score += 0.25
        elif label_ok and formulas_ok >= 4:
            partial = 0.15
            print(f"PARTIAL: Component 1 — Label OK, {formulas_ok}/8 SUM formulas ({partial} pts)")
            total_score += partial
        elif label_ok or formulas_ok > 0:
            partial = 0.05
            print(f"PARTIAL: Component 1 — label_ok={label_ok}, formulas={formulas_ok}/8 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — A52={a52_val}, no SUM formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Bold formatting on rows 1 and 52 (0.15 points)
    try:
        bold_r1 = 0
        bold_r52 = 0
        for c in range(1, 12):
            if ws.cell(row=1, column=c).font.bold:
                bold_r1 += 1
            if ws.cell(row=52, column=c).font.bold:
                bold_r52 += 1

        r1_ok = bold_r1 >= 9  # at least 9 of 11 columns bold
        r52_ok = bold_r52 >= 6  # at least 6 of 11 bold (some may be empty)

        if r1_ok and r52_ok:
            print(f"PASS: Component 2 — Row 1 bold ({bold_r1}/11), Row 52 bold ({bold_r52}/11) (0.15 pts)")
            total_score += 0.15
        elif r1_ok or r52_ok:
            partial = 0.08
            print(f"PARTIAL: Component 2 — Row1 bold={bold_r1}/11 (ok={r1_ok}), Row52 bold={bold_r52}/11 (ok={r52_ok}) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Row1 bold={bold_r1}/11, Row52 bold={bold_r52}/11")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: USD currency number format on D2:K52 (0.20 points)
    # Check $#,##0.00 format on data range
    try:
        currency_count = 0
        total_cells = 0
        for r in range(2, 53):  # rows 2-52
            for c in range(4, 12):  # columns D-K
                total_cells += 1
                nf = ws.cell(row=r, column=c).number_format
                if nf is not None and '$' in str(nf):
                    currency_count += 1

        ratio = currency_count / total_cells if total_cells > 0 else 0
        if ratio >= 0.9:
            print(f"PASS: Component 3 — {currency_count}/{total_cells} cells have USD currency format (0.20 pts)")
            total_score += 0.20
        elif ratio >= 0.5:
            partial = 0.10
            print(f"PARTIAL: Component 3 — {currency_count}/{total_cells} cells ({ratio:.0%}) have currency format ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {currency_count}/{total_cells} cells ({ratio:.0%}) have currency format")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Conditional formatting on K2:K51 — lessThan 2000, red background (0.15 points)
    try:
        # Count how many CF conditions are met: range covers K, rule is <2000, fill is red
        cf_checks_passed = 0

        for cf in ws.conditional_formatting:
            cf_range_str = str(cf).strip()
            # Check if K2:K51 is covered
            if 'K' in cf_range_str:
                cf_checks_passed += 1  # range found
                for rule in cf.rules:
                    if rule.type == 'cellIs' and rule.operator == 'lessThan':
                        if rule.formula and any('2000' in str(f) for f in rule.formula):
                            cf_checks_passed += 1  # correct rule
                            if rule.dxf and rule.dxf.fill:
                                try:
                                    rgb = rule.dxf.fill.fgColor.rgb
                                    if rgb and 'FF0000' in str(rgb):
                                        cf_checks_passed += 1  # red fill
                                except:
                                    pass

        if cf_checks_passed >= 3:
            print(f"PASS: Component 4 — Conditional formatting on K column: <2000 with red fill (0.15 pts)")
            total_score += 0.15
        elif cf_checks_passed >= 2:
            partial = 0.10
            print(f"PARTIAL: Component 4 — CF rule found (<2000) but fill color not red ({partial} pts)")
            total_score += partial
        elif cf_checks_passed >= 1:
            partial = 0.05
            print(f"PARTIAL: Component 4 — CF exists on K but rule not correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — No conditional formatting found on K column")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Thin borders on A1:K52 (0.15 points)
    try:
        bordered_count = 0
        total_border_cells = 0
        # Sample check: check borders on a representative set of cells
        sample_rows = [1, 2, 10, 25, 40, 51, 52]
        sample_cols = [1, 4, 7, 11]  # A, D, G, K
        for r in sample_rows:
            for c in sample_cols:
                total_border_cells += 1
                cell = ws.cell(row=r, column=c)
                border = cell.border
                has_borders = (
                    border.left.style is not None and
                    border.right.style is not None and
                    border.top.style is not None and
                    border.bottom.style is not None
                )
                if has_borders:
                    bordered_count += 1

        ratio = bordered_count / total_border_cells if total_border_cells > 0 else 0
        if ratio >= 0.85:
            print(f"PASS: Component 5 — {bordered_count}/{total_border_cells} sampled cells have borders (0.15 pts)")
            total_score += 0.15
        elif ratio >= 0.5:
            partial = 0.08
            print(f"PARTIAL: Component 5 — {bordered_count}/{total_border_cells} sampled cells have borders ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — Only {bordered_count}/{total_border_cells} sampled cells have borders")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Freeze panes at A2 (row 1 frozen) (0.10 points)
    try:
        freeze = ws.freeze_panes
        if freeze == 'A2':
            print(f"PASS: Component 6 — Freeze panes set to A2 (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 6 — Freeze panes = {freeze}, expected A2")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved LibreOffice changes before verification
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(1.0)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
persist_app_state()
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
