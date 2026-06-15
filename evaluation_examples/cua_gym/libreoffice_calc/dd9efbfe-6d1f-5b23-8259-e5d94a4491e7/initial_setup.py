"""
Initial Setup: Payroll summary spreadsheet with 50 employee rows, raw data only.
Task ID: calc_gsd_044
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_044'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Payroll'

    # Headers in row 1
    headers = [
        'Employee ID', 'Name', 'Department', 'Base Salary', 'Overtime Pay',
        'Gross Pay', 'Federal Tax', 'State Tax', 'Benefits',
        'Total Deductions', 'Net Pay'
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic employee data
    first_names = [
        'Sarah', 'Marcus', 'Priya', 'James', 'Aisha', 'Carlos', 'Emily',
        'David', 'Mei', 'Robert', 'Fatima', 'Alex', 'Yuki', 'Thomas',
        'Zara', 'Kevin', 'Natasha', 'Benjamin', 'Olivia', 'Hassan',
        'Rachel', 'Diego', 'Sophie', 'Michael', 'Leila', 'Brandon',
        'Christina', 'Raj', 'Amanda', 'Tyler', 'Jasmine', 'Patrick',
        'Lisa', 'Derek', 'Monica', 'Samuel', 'Grace', 'Victor',
        'Helen', 'Nathan', 'Diana', 'Frank', 'Julia', 'George',
        'Catherine', 'Andre', 'Michelle', 'Stephen', 'Karen', 'Leo'
    ]
    last_names = [
        'Chen', 'Johnson', 'Patel', 'Williams', 'Ibrahim', 'Rodriguez',
        'Thompson', 'Kim', 'Zhang', 'Martinez', 'Hassan', 'Novak',
        'Tanaka', 'Anderson', 'Okafor', 'Lee', 'Volkov', 'Davis',
        'Parker', 'Ali', 'Goldstein', 'Morales', 'Dupont', 'O\'Brien',
        'Farahani', 'Cooper', 'Reyes', 'Sharma', 'Foster', 'Brooks',
        'Washington', 'Kelly', 'Nguyen', 'Harrison', 'Santos', 'Wright',
        'Liu', 'Fernandez', 'Mitchell', 'Byrne', 'Campbell', 'Romano',
        'Weber', 'Clark', 'Dubois', 'Jackson', 'Rivera', 'Powell',
        'Adams', 'Nakamura'
    ]
    departments = [
        'Engineering', 'Marketing', 'Sales', 'Human Resources', 'Finance',
        'Operations', 'Customer Support', 'Legal', 'Product', 'IT'
    ]

    random.seed(42)  # Reproducible data

    for i in range(50):
        row = i + 2
        emp_id = f'EMP-{1001 + i}'
        name = f'{first_names[i]} {last_names[i]}'
        dept = departments[i % len(departments)]

        base_salary = round(random.uniform(2800, 8500), 2)
        overtime = round(random.uniform(0, 1500), 2)
        gross_pay = round(base_salary + overtime, 2)

        federal_tax = round(gross_pay * random.uniform(0.10, 0.22), 2)
        state_tax = round(gross_pay * random.uniform(0.03, 0.08), 2)
        benefits = round(random.uniform(150, 600), 2)

        total_deductions = round(federal_tax + state_tax + benefits, 2)
        net_pay = round(gross_pay - total_deductions, 2)

        ws.cell(row=row, column=1, value=emp_id)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=dept)
        ws.cell(row=row, column=4, value=base_salary)
        ws.cell(row=row, column=5, value=overtime)
        ws.cell(row=row, column=6, value=gross_pay)
        ws.cell(row=row, column=7, value=federal_tax)
        ws.cell(row=row, column=8, value=state_tax)
        ws.cell(row=row, column=9, value=benefits)
        ws.cell(row=row, column=10, value=total_deductions)
        ws.cell(row=row, column=11, value=net_pay)

    # Row 52 is intentionally empty (no totals yet)
    # No formatting, no formulas, no borders, no freeze panes

    # Set reasonable column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 18
    for col_letter in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        ws.column_dimensions[col_letter].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
