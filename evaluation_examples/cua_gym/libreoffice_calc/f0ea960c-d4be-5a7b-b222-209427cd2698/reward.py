"""
Reward Script: AutoFill VLOOKUP formula from B2 down to B40
Task ID: calc_cop_autofill_007
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: B3:B40 filled with VLOOKUP formulas (0.5 pts)
               - Initial state: B3:B40 are empty (None)
               - Golden state: B3:B40 contain =VLOOKUP(An,$G$2:$H$50,2,0)
  Component 2: All formulas in B3:B40 use absolute reference $G$2:$H$50 (0.3 pts)
               - Verifies the lookup table reference stays fixed during autofill
  Component 3: All formulas in B3:B40 use correct relative row reference A[n] (0.2 pts)
               - Verifies the lookup_value updates for each row (relative reference)
Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_cop_autofill_007'
SHEET_NAME = 'Lookup'

# Expected formula pattern:
# Each row n (3..40) should have =VLOOKUP(An,$G$2:$H$50,2,0)
# The lookup table $G$2:$H$50 must be absolute, An must be relative
EXPECTED_ABSOLUTE_REF = '$G$2:$H$50'
EXPECTED_COL_INDEX = '2'
EXPECTED_MATCH_TYPE = '0'
AUTOFILL_START_ROW = 3
AUTOFILL_END_ROW = 40


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: sheet 'Lookup' must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found in workbook. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Component 1: B3:B40 are filled with VLOOKUP formulas (0.5 points)
    # In the initial file, B3:B40 are all None. In the golden file, they contain formulas.
    try:
        filled_count = 0
        formula_cells = []
        empty_cells = []

        for row in range(AUTOFILL_START_ROW, AUTOFILL_END_ROW + 1):
            cell_val = ws.cell(row=row, column=2).value
            if cell_val is not None and isinstance(cell_val, str) and cell_val.upper().startswith('=VLOOKUP('):
                filled_count += 1
                formula_cells.append(row)
            else:
                empty_cells.append(row)

        expected_count = AUTOFILL_END_ROW - AUTOFILL_START_ROW + 1  # 38 cells
        if filled_count == expected_count:
            print(f"PASS: Component 1 — All {expected_count} cells B{AUTOFILL_START_ROW}:B{AUTOFILL_END_ROW} contain VLOOKUP formulas (0.5 pts)")
            total_score += 0.5
        elif filled_count > 0:
            partial = round(0.5 * filled_count / expected_count, 3)
            print(f"PARTIAL: Component 1 — {filled_count}/{expected_count} cells have VLOOKUP formulas. Empty rows: {empty_cells[:5]}{'...' if len(empty_cells) > 5 else ''} ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No VLOOKUP formulas found in B{AUTOFILL_START_ROW}:B{AUTOFILL_END_ROW}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All formulas use the correct absolute reference $G$2:$H$50 (0.3 points)
    # This verifies the lookup table reference stays fixed (absolute) during AutoFill.
    # Only checks rows that had formulas (from component 1).
    try:
        abs_ref_correct = 0
        abs_ref_wrong = []

        for row in range(AUTOFILL_START_ROW, AUTOFILL_END_ROW + 1):
            cell_val = ws.cell(row=row, column=2).value
            if cell_val is not None and isinstance(cell_val, str) and cell_val.upper().startswith('=VLOOKUP('):
                # Normalize: uppercase and remove spaces
                normalized = cell_val.upper().replace(' ', '')
                if EXPECTED_ABSOLUTE_REF in normalized:
                    abs_ref_correct += 1
                else:
                    abs_ref_wrong.append((row, cell_val))

        if formula_cells:  # only score if there were formulas to check
            if abs_ref_correct == len(formula_cells) and len(abs_ref_wrong) == 0:
                print(f"PASS: Component 2 — All {abs_ref_correct} formulas use absolute reference {EXPECTED_ABSOLUTE_REF} (0.3 pts)")
                total_score += 0.3
            elif abs_ref_correct > 0:
                partial = round(0.3 * abs_ref_correct / len(formula_cells), 3)
                print(f"PARTIAL: Component 2 — {abs_ref_correct}/{len(formula_cells)} formulas use correct absolute ref. Wrong: {abs_ref_wrong[:3]} ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 2 — No formulas use the correct absolute reference {EXPECTED_ABSOLUTE_REF}. Examples: {abs_ref_wrong[:3]}")
        else:
            print("SKIP: Component 2 — No formulas found to check absolute reference")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Each formula uses the correct relative row reference A[n] (0.2 points)
    # =VLOOKUP(A3,...) in B3, =VLOOKUP(A4,...) in B4, etc.
    # This verifies the lookup_value updates properly for each row (relative reference behavior).
    try:
        relative_ref_correct = 0
        relative_ref_wrong = []

        for row in range(AUTOFILL_START_ROW, AUTOFILL_END_ROW + 1):
            cell_val = ws.cell(row=row, column=2).value
            if cell_val is not None and isinstance(cell_val, str) and cell_val.upper().startswith('=VLOOKUP('):
                # Normalize to uppercase, remove spaces
                normalized = cell_val.upper().replace(' ', '')
                # Expected: =VLOOKUP(A{row},$G$2:$H$50,2,0)
                expected_lookup_val = f'A{row}'
                # Extract first argument of VLOOKUP
                # Pattern: =VLOOKUP(A3,$G$2:$H$50,2,0)
                match = re.match(r'=VLOOKUP\(([^,]+),', normalized)
                if match:
                    actual_lookup_val = match.group(1).strip()
                    if actual_lookup_val == expected_lookup_val:
                        relative_ref_correct += 1
                    else:
                        relative_ref_wrong.append((row, cell_val, f"expected {expected_lookup_val}, got {actual_lookup_val}"))
                else:
                    relative_ref_wrong.append((row, cell_val, "could not parse lookup value"))

        if formula_cells:  # only score if there were formulas to check
            if relative_ref_correct == len(formula_cells) and len(relative_ref_wrong) == 0:
                print(f"PASS: Component 3 — All {relative_ref_correct} formulas use correct relative row reference A[n] (0.2 pts)")
                total_score += 0.2
            elif relative_ref_correct > 0:
                partial = round(0.2 * relative_ref_correct / len(formula_cells), 3)
                print(f"PARTIAL: Component 3 — {relative_ref_correct}/{len(formula_cells)} formulas use correct relative ref. Wrong: {relative_ref_wrong[:3]} ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 3 — No formulas have correct relative row reference. Examples: {relative_ref_wrong[:3]}")
        else:
            print("SKIP: Component 3 — No formulas found to check relative reference")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 3), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
