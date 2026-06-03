"""
Initial Setup: AutoFill VLOOKUP formula from B2 down to B40
Task ID: calc_cop_autofill_007
Domain: libreoffice_calc

Creates a spreadsheet with:
  - Sheet 'Lookup':
    - A2:A40: IDs to look up
    - G2:H50: Lookup table (G=ID, H=Name)
    - B2: =VLOOKUP(A2,$G$2:$H$50,2,0) (already present)
    - B3:B40: EMPTY (agent must autofill)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_cop_autofill_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Lookup'

    # --- Column headers (row 1) ---
    ws['A1'] = 'Employee ID'
    ws['B1'] = 'Employee Name'
    ws['G1'] = 'ID'
    ws['H1'] = 'Name'

    # Style headers bold
    for cell in ['A1', 'B1', 'G1', 'H1']:
        ws[cell].font = Font(bold=True)

    # --- Lookup table: G2:H50 (49 entries) ---
    lookup_data = [
        ('E001', 'Sarah Chen'),
        ('E002', 'Marcus Johnson'),
        ('E003', 'Priya Patel'),
        ('E004', 'David Kim'),
        ('E005', 'Emily Nguyen'),
        ('E006', 'James Wilson'),
        ('E007', 'Olivia Martinez'),
        ('E008', 'Liam Thompson'),
        ('E009', 'Ava Robinson'),
        ('E010', 'Noah Garcia'),
        ('E011', 'Isabella Brown'),
        ('E012', 'Mason Lee'),
        ('E013', 'Sophia White'),
        ('E014', 'Ethan Harris'),
        ('E015', 'Mia Clark'),
        ('E016', 'Benjamin Lewis'),
        ('E017', 'Charlotte Walker'),
        ('E018', 'William Hall'),
        ('E019', 'Amelia Allen'),
        ('E020', 'Lucas Young'),
        ('E021', 'Harper King'),
        ('E022', 'Henry Scott'),
        ('E023', 'Ella Adams'),
        ('E024', 'Alexander Baker'),
        ('E025', 'Chloe Nelson'),
        ('E026', 'Daniel Carter'),
        ('E027', 'Lily Mitchell'),
        ('E028', 'Matthew Perez'),
        ('E029', 'Grace Roberts'),
        ('E030', 'Sebastian Turner'),
        ('E031', 'Zoey Phillips'),
        ('E032', 'Jackson Campbell'),
        ('E033', 'Penelope Parker'),
        ('E034', 'Aiden Evans'),
        ('E035', 'Riley Edwards'),
        ('E036', 'Owen Collins'),
        ('E037', 'Layla Stewart'),
        ('E038', 'Gabriel Sanchez'),
        ('E039', 'Nora Morris'),
        ('E040', 'Ryan Rogers'),
        ('E041', 'Zoe Reed'),
        ('E042', 'Christian Cook'),
        ('E043', 'Naomi Morgan'),
        ('E044', 'Elijah Bell'),
        ('E045', 'Aurora Murphy'),
        ('E046', 'Aaron Bailey'),
        ('E047', 'Violet Rivera'),
        ('E048', 'Charles Cooper'),
        ('E049', 'Stella Richardson'),
    ]

    for i, (emp_id, name) in enumerate(lookup_data, start=2):
        ws.cell(row=i, column=7, value=emp_id)   # G column
        ws.cell(row=i, column=8, value=name)     # H column

    # --- A2:A40: IDs to look up (shuffled subset from the lookup table) ---
    lookup_ids = [
        'E003', 'E017', 'E001', 'E042', 'E025',
        'E008', 'E034', 'E019', 'E047', 'E012',
        'E028', 'E006', 'E039', 'E015', 'E022',
        'E049', 'E031', 'E010', 'E043', 'E007',
        'E016', 'E036', 'E004', 'E045', 'E023',
        'E011', 'E038', 'E029', 'E005', 'E020',
        'E048', 'E032', 'E013', 'E040', 'E021',
        'E002', 'E046', 'E033', 'E014',
    ]
    # A2:A40 -> 39 rows
    for i, emp_id in enumerate(lookup_ids, start=2):
        ws.cell(row=i, column=1, value=emp_id)

    # --- B2: VLOOKUP formula (already present) ---
    ws['B2'] = '=VLOOKUP(A2,$G$2:$H$50,2,0)'

    # B3:B40 remain EMPTY (agent must autofill)

    # Column widths for readability
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 20

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

create_initial()
