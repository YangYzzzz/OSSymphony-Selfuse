"""
Initial Setup: Department budget spreadsheet for absolute reference formula task
Task ID: calc_fmb_absolute_ref_041
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_absolute_ref_041'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Department Budgets ---
    ws = wb.active
    ws.title = 'Department Budgets'

    # Headers
    ws['A1'] = 'Department'
    ws['B1'] = 'Budget'
    ws['C1'] = '% of Total'

    # Department data
    data = [
        ('Engineering', 450000),
        ('Marketing',   180000),
        ('Sales',       220000),
        ('Finance',      95000),
        ('HR',           75000),
        ('IT',          130000),
        ('Operations',  160000),
        ('Legal',        85000),
        ('Procurement',  70000),
        ('Admin',        55000),
    ]

    for r, (dept, budget) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=dept)
        ws.cell(row=r, column=2, value=budget)
        # C column (% of Total) is intentionally left empty — the agent must fill C2

    # TOTAL row (row 12)
    ws['A12'] = 'TOTAL'
    ws['B12'] = 1520000   # sum of all departments above

    # C2 is intentionally EMPTY — this is the cell the agent needs to fill

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
