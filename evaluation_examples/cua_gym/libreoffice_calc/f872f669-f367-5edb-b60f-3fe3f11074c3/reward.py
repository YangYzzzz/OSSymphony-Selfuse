"""
Reward Script: Enter a percentage formula in C2 with absolute reference to B12
Task ID: calc_fmb_absolute_ref_041
Domain: libreoffice_calc
Scoring:
  Component 1: C2 contains a formula referencing B2 and B12 — 0.5 pts
  Component 2: The formula uses a fully absolute reference ($B$12) for B12 — 0.5 pts
Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmb_absolute_ref_041'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task: In cell C2 of 'Department Budgets' sheet, enter a formula that calculates
    what percentage each department's budget (column B) represents of the total budget
    in cell B12. Use an absolute reference for B12 ($B$12).

    Initial state: C2 is empty.
    Golden state: C2 contains =B2/$B$12 (or equivalent formula with $B$12).
    """
    total_score = 0.0

    # Precondition gate: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: 'Department Budgets' sheet must exist
    if 'Department Budgets' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Department Budgets' not found.")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Department Budgets']

    # Component 1: C2 contains a non-empty formula (starts with '=') — 0.5 points
    # CHANGES between initial (None) and golden (=B2/$B$12): this check FAILS on initial, PASSES on golden.
    try:
        c2_value = ws['C2'].value
        if c2_value is not None and isinstance(c2_value, str) and c2_value.strip().startswith('='):
            print(f"PASS: Component 1 — C2 contains a formula: {repr(c2_value)} (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — C2 does not contain a formula; found: {repr(c2_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — could not read C2: {e}")

    # Component 2: The formula in C2 uses a fully absolute reference $B$12 — 0.5 points
    # The task specifically requires absolute reference so formula can be safely copied down.
    # $B$12 locks both column B and row 12.
    # CHANGES between initial (empty) and golden (=B2/$B$12): FAILS on initial, PASSES on golden.
    try:
        c2_value = ws['C2'].value
        if (c2_value is not None
                and isinstance(c2_value, str)
                and c2_value.strip().startswith('=')):
            formula_upper = c2_value.upper().replace(' ', '')
            # Must contain $B$12 — absolute column AND absolute row reference
            if '$B$12' in formula_upper:
                print(f"PASS: Component 2 — Formula uses absolute reference $B$12: {repr(c2_value)} (0.5 pts)")
                total_score += 0.5
            else:
                # Provide diagnostic: partial absolute or no absolute
                if re.search(r'[\$]?B[\$]?12', formula_upper):
                    print(f"FAIL: Component 2 — Formula references B12 but not with full absolute $B$12: {repr(c2_value)}")
                else:
                    print(f"FAIL: Component 2 — Formula does not reference $B$12 at all: {repr(c2_value)}")
        else:
            print(f"FAIL: Component 2 — C2 has no formula; cannot check absolute reference.")
    except Exception as e:
        print(f"ERROR: Component 2 — could not check absolute reference in C2: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
