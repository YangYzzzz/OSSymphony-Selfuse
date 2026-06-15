"""
Initial Setup: Format cells B2:B20 as accounting format
Task ID: calc_gfl_062
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_062'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: GL (General Ledger) ---
    ws = wb.active
    ws.title = 'GL'

    # Headers
    headers = ['Account', 'Amount', 'Debit', 'Credit', 'Balance']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 19 rows of realistic general ledger data (rows 2-20)
    ledger_data = [
        ['1010 - Cash',                  15000.00,  15000.00,      0.00,  15000.00],
        ['1020 - Accounts Receivable',    2500.50,   2500.50,      0.00,  17500.50],
        ['2010 - Accounts Payable',      -8000.00,      0.00,   8000.00,   9500.50],
        ['4010 - Sales Revenue',         32750.00,      0.00,  32750.00, -23249.50],
        ['5010 - Cost of Goods Sold',    18400.00,  18400.00,      0.00,  -4849.50],
        ['1010 - Cash',                   4200.75,   4200.75,      0.00,   -648.75],
        ['6010 - Rent Expense',           3500.00,   3500.00,      0.00,   2851.25],
        ['6020 - Utilities Expense',       875.30,    875.30,      0.00,   3726.55],
        ['1030 - Inventory',            -12500.00,      0.00,  12500.00,  -8773.45],
        ['2020 - Notes Payable',         -5000.00,      0.00,   5000.00, -13773.45],
        ['4010 - Sales Revenue',         21680.00,      0.00,  21680.00, -35453.45],
        ['6030 - Salaries Expense',      14200.00,  14200.00,      0.00, -21253.45],
        ['1010 - Cash',                   9350.25,   9350.25,      0.00, -11903.20],
        ['6040 - Office Supplies',         430.00,    430.00,      0.00, -11473.20],
        ['2010 - Accounts Payable',      -3200.00,      0.00,   3200.00, -14673.20],
        ['1020 - Accounts Receivable',    6780.50,   6780.50,      0.00,  -7892.70],
        ['6050 - Insurance Expense',      1950.00,   1950.00,      0.00,  -5942.70],
        ['4020 - Service Revenue',        8500.00,      0.00,   8500.00, -14442.70],
        ['6060 - Depreciation Expense',   2100.00,   2100.00,      0.00, -12342.70],
    ]

    for r, row_data in enumerate(ledger_data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

    # Column B has plain number format (General) - NO accounting format
    # This is the default, but be explicit
    for row in range(2, 21):
        ws.cell(row=row, column=2).number_format = 'General'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
