"""
Reward Script: Format cells B2:B20 as accounting format
Task ID: calc_gfl_062
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): All 19 cells have accounting number format (contains $ and parentheses pattern)
  Component 2 (0.3): Exact standard accounting format string matches
  Component 3 (0.3): Data values preserved correctly after formatting
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_062'

# Expected accounting format string (standard Excel/Calc accounting format)
ACCOUNTING_FORMAT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

# Expected values in B2:B20 (from the initial data, should be unchanged)
EXPECTED_VALUES = {
    2: 15000, 3: 2500.5, 4: -8000, 5: 32750, 6: 18400,
    7: 4200.75, 8: 3500, 9: 875.3, 10: -12500, 11: -5000,
    12: 21680, 13: 14200, 14: 9350.25, 15: 430, 16: -3200,
    17: 6780.5, 18: 1950, 19: 8500, 20: 2100,
}


def persist_app_state(domain):
    """Save any unsaved changes in LibreOffice before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify that cells B2:B20 have accounting format applied.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    try:
        ws = wb['GL']
    except KeyError:
        print("CRITICAL: Sheet 'GL' not found")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: All 19 cells B2:B20 have a non-General accounting-style format (0.4 points)
    # An accounting format must contain '$' and use parentheses for negatives
    try:
        accounting_count = 0
        for r in range(2, 21):
            nf = ws.cell(row=r, column=2).number_format
            if nf and nf != 'General' and '$' in nf and '(' in nf:
                accounting_count += 1
            else:
                print(f"  DETAIL: B{r} format is {nf!r} — not accounting")

        if accounting_count == 19:
            print(f"PASS: Component 1 — All 19 cells have accounting format (0.4 pts)")
            total_score += 0.4
        elif accounting_count > 0:
            partial = round(0.4 * (accounting_count / 19), 2)
            print(f"FAIL: Component 1 — {accounting_count}/19 cells have accounting format (partial {partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No cells have accounting format (0/19)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Exact standard accounting format string (0.3 points)
    # Checks that the format matches the standard accounting pattern precisely
    try:
        exact_count = 0
        for r in range(2, 21):
            nf = ws.cell(row=r, column=2).number_format
            if nf == ACCOUNTING_FORMAT:
                exact_count += 1

        if exact_count == 19:
            print(f"PASS: Component 2 — All 19 cells have exact standard accounting format (0.3 pts)")
            total_score += 0.3
        elif exact_count > 0:
            # Accept alternative accounting formats that still have $ and parentheses
            # Check if all cells at least have SOME accounting format (from Component 1)
            # Give partial credit proportional to exact matches
            partial = round(0.3 * (exact_count / 19), 2)
            print(f"FAIL: Component 2 — {exact_count}/19 cells have exact format (partial {partial} pts)")
            total_score += partial
        else:
            # Check if they have an alternative accounting format (still award some credit)
            alt_count = 0
            for r in range(2, 21):
                nf = ws.cell(row=r, column=2).number_format
                if nf and nf != 'General' and '$' in nf and '#' in nf and '0.00' in nf:
                    alt_count += 1
            if alt_count == 19:
                print(f"PARTIAL: Component 2 — All cells have alternative accounting format (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 2 — No cells match standard accounting format")
                sample_nf = ws.cell(row=2, column=2).number_format
                print(f"  DETAIL: B2 format is {sample_nf!r}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Data values preserved correctly (0.3 points)
    # Formatting should not alter the underlying numeric values
    # This component only awards points if the format has ALSO changed (compound check)
    try:
        format_changed = ws.cell(row=2, column=2).number_format != 'General'
        if not format_changed:
            print(f"FAIL: Component 3 — Format not changed, so data preservation check is moot (0 pts)")
        else:
            preserved_count = 0
            for r in range(2, 21):
                cell_val = ws.cell(row=r, column=2).value
                expected = EXPECTED_VALUES[r]
                if cell_val is not None:
                    try:
                        if abs(float(cell_val) - expected) < 0.01:
                            preserved_count += 1
                        else:
                            print(f"  DETAIL: B{r} value {cell_val} != expected {expected}")
                    except (ValueError, TypeError):
                        print(f"  DETAIL: B{r} value {cell_val!r} is not numeric")
                else:
                    print(f"  DETAIL: B{r} is None")

            if preserved_count == 19:
                print(f"PASS: Component 3 — All 19 values preserved after formatting (0.3 pts)")
                total_score += 0.3
            elif preserved_count > 0:
                partial = round(0.3 * (preserved_count / 19), 2)
                print(f"FAIL: Component 3 — {preserved_count}/19 values preserved (partial {partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 3 — No values preserved correctly")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
