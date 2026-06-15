"""
Initial Setup: Configure sheet to print row and column headers
Task ID: calc_gfl_093
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_093'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Reference ---
    ws = wb.active
    ws.title = 'Reference'

    # Headers (8 columns)
    headers = [
        'Product ID', 'Product Name', 'Category', 'Supplier',
        'Unit Price', 'Stock Qty', 'Reorder Level', 'Last Updated'
    ]
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows (25 rows of realistic product catalog data)
    data = [
        ['PRD-1001', 'Wireless Bluetooth Headphones', 'Electronics', 'TechVision Inc.', 79.99, 145, 30, '2025-11-03'],
        ['PRD-1002', 'Organic Green Tea (100 bags)', 'Beverages', 'Mountain Leaf Co.', 12.50, 320, 50, '2025-10-28'],
        ['PRD-1003', 'Stainless Steel Water Bottle', 'Kitchen', 'EcoLife Products', 24.95, 210, 40, '2025-11-01'],
        ['PRD-1004', 'Ergonomic Office Chair', 'Furniture', 'ComfortPlus Ltd.', 349.00, 42, 10, '2025-09-15'],
        ['PRD-1005', 'LED Desk Lamp with USB Port', 'Electronics', 'BrightSpace Co.', 45.99, 88, 20, '2025-10-22'],
        ['PRD-1006', 'Premium Notebook Set (3-pack)', 'Stationery', 'PaperCraft Inc.', 18.75, 415, 75, '2025-11-05'],
        ['PRD-1007', 'Bamboo Cutting Board Set', 'Kitchen', 'EcoLife Products', 32.00, 167, 35, '2025-10-30'],
        ['PRD-1008', 'Portable External SSD 1TB', 'Electronics', 'TechVision Inc.', 89.99, 73, 15, '2025-11-02'],
        ['PRD-1009', 'Yoga Mat with Carry Strap', 'Sports', 'ActiveWear Global', 28.50, 192, 45, '2025-10-18'],
        ['PRD-1010', 'French Press Coffee Maker', 'Kitchen', 'BrewMaster Co.', 34.99, 128, 25, '2025-10-25'],
        ['PRD-1011', 'Adjustable Standing Desk', 'Furniture', 'ComfortPlus Ltd.', 499.00, 31, 8, '2025-09-20'],
        ['PRD-1012', 'Wireless Charging Pad', 'Electronics', 'TechVision Inc.', 19.99, 256, 60, '2025-11-04'],
        ['PRD-1013', 'Ceramic Plant Pot Set (4-pack)', 'Home Decor', 'GreenThumb Supply', 42.00, 94, 20, '2025-10-12'],
        ['PRD-1014', 'Insulated Lunch Bag', 'Kitchen', 'EcoLife Products', 22.50, 178, 40, '2025-10-29'],
        ['PRD-1015', 'Noise Cancelling Earbuds', 'Electronics', 'BrightSpace Co.', 129.99, 56, 12, '2025-11-01'],
        ['PRD-1016', 'Resistance Band Set (5-pack)', 'Sports', 'ActiveWear Global', 15.99, 340, 70, '2025-10-20'],
        ['PRD-1017', 'Scented Candle Collection', 'Home Decor', 'Aromatic Bliss', 38.00, 112, 25, '2025-10-15'],
        ['PRD-1018', 'USB-C Hub 7-in-1', 'Electronics', 'TechVision Inc.', 54.99, 98, 20, '2025-11-03'],
        ['PRD-1019', 'Reusable Grocery Bags (6-pack)', 'Kitchen', 'EcoLife Products', 14.99, 520, 100, '2025-10-27'],
        ['PRD-1020', 'Desk Organizer with Drawers', 'Stationery', 'PaperCraft Inc.', 29.50, 143, 30, '2025-10-31'],
        ['PRD-1021', 'Foam Roller for Recovery', 'Sports', 'ActiveWear Global', 21.99, 205, 50, '2025-10-19'],
        ['PRD-1022', 'Smart Power Strip (6 outlets)', 'Electronics', 'BrightSpace Co.', 39.99, 77, 15, '2025-11-02'],
        ['PRD-1023', 'Linen Throw Blanket', 'Home Decor', 'CozyHome Textiles', 55.00, 63, 12, '2025-10-08'],
        ['PRD-1024', 'Glass Food Storage Set', 'Kitchen', 'EcoLife Products', 44.95, 89, 18, '2025-10-26'],
        ['PRD-1025', 'Mechanical Keyboard (RGB)', 'Electronics', 'TechVision Inc.', 74.99, 115, 25, '2025-11-05'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 5:  # Unit Price
                cell.number_format = '$#,##0.00'
            elif c == 6 or c == 7:  # Stock Qty, Reorder Level
                cell.number_format = '#,##0'

    # Set column widths for readability
    col_widths = {'A': 12, 'B': 35, 'C': 14, 'D': 22, 'E': 12, 'F': 10, 'G': 14, 'H': 14}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Ensure print headings is NOT set (default state - task requires enabling it)
    ws.print_options.headings = False

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
