"""
Reward Script: Configure sheet to print row and column headers
Task ID: calc_gfl_093
Domain: libreoffice_calc
Scoring:
  Component 1 (0.6): print_options.headings is True on 'Reference' sheet
  Component 2 (0.4): headings enabled AND data integrity preserved (sheet structure intact)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_093'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Reference' sheet must exist
    if 'Reference' not in wb.sheetnames:
        print(f"FAIL: 'Reference' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Reference']

    # Component 1: print_options.headings is True (0.6 points)
    # This is the core task requirement — enable row/column headers for printing.
    # In openpyxl, this maps to ws.print_options.headings.
    try:
        headings_enabled = ws.print_options.headings
        if headings_enabled is True:
            print(f"PASS: Component 1 — print_options.headings is True (0.6 pts)")
            total_score += 0.6
        else:
            print(f"FAIL: Component 1 — print_options.headings is {headings_enabled}, expected True")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: headings enabled AND data integrity preserved (0.4 points)
    # Compound check: the setting must be on AND the spreadsheet data must be intact.
    # This ensures the agent didn't corrupt data while changing the print setting.
    try:
        headings_ok = ws.print_options.headings is True
        if not headings_ok:
            print(f"FAIL: Component 2 — headings not enabled, skipping data integrity check")
        else:
            # Verify sheet structure: 25 data rows + 1 header = 26 rows, 8 columns
            data_ok = True
            issues = []

            # Check dimensions
            if ws.max_row < 26:
                data_ok = False
                issues.append(f"max_row={ws.max_row}, expected >=26")
            if ws.max_column < 8:
                data_ok = False
                issues.append(f"max_column={ws.max_column}, expected >=8")

            # Spot-check header row
            expected_headers = ['Product ID', 'Product Name', 'Category', 'Supplier',
                                'Unit Price', 'Stock Qty', 'Reorder Level', 'Last Updated']
            actual_headers = [ws.cell(row=1, column=c).value for c in range(1, 9)]
            if actual_headers != expected_headers:
                data_ok = False
                issues.append(f"headers mismatch: {actual_headers}")

            # Spot-check a known data cell: B2 should be 'Wireless Bluetooth Headphones'
            b2_val = ws.cell(row=2, column=2).value
            if b2_val != 'Wireless Bluetooth Headphones':
                data_ok = False
                issues.append(f"B2 value={b2_val}, expected 'Wireless Bluetooth Headphones'")

            # Spot-check last row: A26 should be 'PRD-1025'
            a26_val = ws.cell(row=26, column=1).value
            if a26_val != 'PRD-1025':
                data_ok = False
                issues.append(f"A26 value={a26_val}, expected 'PRD-1025'")

            if data_ok:
                print(f"PASS: Component 2 — headings enabled AND data integrity verified (0.4 pts)")
                total_score += 0.4
            else:
                print(f"FAIL: Component 2 — data integrity issues: {'; '.join(issues)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
