"""
Reward Script: ETL transformation pipeline via Macros on InputData sheet
Task ID: calc_gg5_050
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25) - TransformedData sheet exists with correct headers
  Component 2 (0.25) - No 'Cancelled' rows in TransformedData
  Component 3 (0.25) - Column K 'Profit Margin' computed as (Revenue-Cost)/Revenue
  Component 4 (0.25) - Data sorted by Date (column B) ascending
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_050'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: InputData sheet must exist
    if 'InputData' not in wb.sheetnames:
        print("FAIL: InputData sheet not found — file may be corrupted")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: TransformedData sheet exists with correct headers (0.25 points)
    try:
        if 'TransformedData' not in wb.sheetnames:
            print("FAIL: Component 1 — 'TransformedData' sheet does not exist")
        else:
            ws_td = wb['TransformedData']
            # Check headers: original 10 columns + 'Profit Margin' in K
            expected_headers = ['OrderID', 'Date', 'Status', 'CustomerID', 'Product',
                                'Category', 'Units', 'Revenue', 'Cost', 'Country', 'Profit Margin']
            actual_headers = [ws_td.cell(row=1, column=c).value for c in range(1, 12)]
            if actual_headers == expected_headers:
                print(f"PASS: Component 1 — TransformedData exists with correct 11 headers (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 1 — Headers mismatch. Expected: {expected_headers}, Got: {actual_headers}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: No 'Cancelled' rows in TransformedData (0.25 points)
    try:
        if 'TransformedData' not in wb.sheetnames:
            print("FAIL: Component 2 — TransformedData sheet missing")
        else:
            ws_td = wb['TransformedData']
            cancelled_count = 0
            data_rows = ws_td.max_row - 1  # exclude header
            for r in range(2, ws_td.max_row + 1):
                status_val = ws_td.cell(row=r, column=3).value
                if status_val is not None and str(status_val).strip() == 'Cancelled':
                    cancelled_count += 1

            if cancelled_count == 0 and data_rows > 0:
                # Also verify row count is reasonable (should be ~350 from 400 - 50 cancelled)
                if 300 <= data_rows <= 400:
                    print(f"PASS: Component 2 — No cancelled rows found. {data_rows} data rows (0.25 pts)")
                    total_score += 0.25
                else:
                    print(f"FAIL: Component 2 — No cancelled rows but unexpected row count: {data_rows} (expected ~350)")
            else:
                print(f"FAIL: Component 2 — Found {cancelled_count} cancelled rows in TransformedData")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Column K 'Profit Margin' correctly computed as (Revenue-Cost)/Revenue (0.25 points)
    try:
        if 'TransformedData' not in wb.sheetnames:
            print("FAIL: Component 3 — TransformedData sheet missing")
        else:
            ws_td = wb['TransformedData']
            if ws_td.max_row < 2:
                print("FAIL: Component 3 — No data rows in TransformedData")
            else:
                correct_count = 0
                checked_count = 0
                # Sample up to 20 rows spread across the data
                sample_rows = list(range(2, min(ws_td.max_row + 1, 22)))
                # Also include some rows from the middle and end
                mid = (ws_td.max_row + 2) // 2
                end = ws_td.max_row
                for r in [mid, mid + 1, end - 1, end]:
                    if 2 <= r <= ws_td.max_row and r not in sample_rows:
                        sample_rows.append(r)

                for r in sample_rows:
                    revenue = ws_td.cell(row=r, column=8).value
                    cost = ws_td.cell(row=r, column=9).value
                    pm = ws_td.cell(row=r, column=11).value
                    if revenue is not None and cost is not None and pm is not None:
                        checked_count += 1
                        try:
                            rev_f = float(revenue)
                            cost_f = float(cost)
                            pm_f = float(pm)
                            if rev_f != 0:
                                expected_pm = (rev_f - cost_f) / rev_f
                                if abs(pm_f - expected_pm) < 0.001:
                                    correct_count += 1
                        except (ValueError, TypeError):
                            pass

                if checked_count > 0 and correct_count == checked_count:
                    print(f"PASS: Component 3 — Profit Margin correctly computed in all {checked_count} sampled rows (0.25 pts)")
                    total_score += 0.25
                elif checked_count > 0:
                    print(f"FAIL: Component 3 — {correct_count}/{checked_count} sampled rows have correct Profit Margin")
                else:
                    print("FAIL: Component 3 — Could not verify Profit Margin values (no numeric data found)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Data sorted by Date (column B) ascending (0.25 points)
    try:
        if 'TransformedData' not in wb.sheetnames:
            print("FAIL: Component 4 — TransformedData sheet missing")
        else:
            ws_td = wb['TransformedData']
            if ws_td.max_row < 3:
                print("FAIL: Component 4 — Not enough data rows to verify sorting")
            else:
                dates = []
                for r in range(2, ws_td.max_row + 1):
                    d = ws_td.cell(row=r, column=2).value
                    dates.append(d)

                # Check all dates are present and sorted ascending
                non_none_dates = [d for d in dates if d is not None]
                if len(non_none_dates) < len(dates) * 0.9:
                    print(f"FAIL: Component 4 — Too many missing dates ({len(dates) - len(non_none_dates)} of {len(dates)})")
                else:
                    is_sorted = all(
                        non_none_dates[i] <= non_none_dates[i + 1]
                        for i in range(len(non_none_dates) - 1)
                    )
                    if is_sorted:
                        print(f"PASS: Component 4 — Data sorted by Date ascending ({len(non_none_dates)} dates verified) (0.25 pts)")
                        total_score += 0.25
                    else:
                        # Find first out-of-order pair for debugging
                        for i in range(len(non_none_dates) - 1):
                            if non_none_dates[i] > non_none_dates[i + 1]:
                                print(f"FAIL: Component 4 — Dates not sorted. Row {i+2}: {non_none_dates[i]} > Row {i+3}: {non_none_dates[i+1]}")
                                break
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
