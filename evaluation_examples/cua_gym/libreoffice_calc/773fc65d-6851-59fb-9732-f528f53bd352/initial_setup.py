"""
Initial Setup: Create InputData sheet with 400 rows for ETL pipeline task
Task ID: calc_gg5_050
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time
from datetime import datetime, timedelta

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_050'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'InputData'

    # Headers: A-J
    headers = [
        'OrderID', 'Date', 'Status', 'CustomerID', 'Product',
        'Category', 'Units', 'Revenue', 'Cost', 'Country'
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Data pools
    statuses_normal = ['Shipped', 'Delivered', 'Processing', 'Pending', 'Returned']
    products = [
        'Laptop Pro 15', 'Wireless Mouse MX', 'USB-C Hub 7-Port', 'Mechanical Keyboard K7',
        'Monitor 27" 4K', '256GB SSD Drive', 'Webcam HD 1080p', 'Bluetooth Speaker',
        'Noise Cancelling Headphones', 'Tablet Stand Adjustable', 'External Battery 20000mAh',
        'HDMI Cable 2m', 'Desk Lamp LED', 'Ergonomic Chair Base', 'Wireless Charger Pad',
        'Graphics Tablet A4', 'Portable Projector', 'Smart Power Strip', 'Cable Management Kit',
        'Screen Protector Pack'
    ]
    categories = ['Electronics', 'Accessories', 'Peripherals', 'Furniture', 'Storage', 'Audio', 'Display']
    countries = [
        'United States', 'Canada', 'United Kingdom', 'Germany', 'France',
        'Japan', 'Australia', 'Brazil', 'India', 'Mexico',
        'South Korea', 'Netherlands', 'Spain', 'Italy', 'Sweden'
    ]
    first_names = [
        'Sarah', 'Marcus', 'Aisha', 'James', 'Yuki', 'Elena', 'Carlos',
        'Priya', 'Oliver', 'Fatima', 'David', 'Lin', 'Sofia', 'Ahmed',
        'Maria', 'Wei', 'Anna', 'Raj', 'Isabella', 'Thomas'
    ]

    base_date = datetime(2024, 1, 5)

    # Decide which rows are Cancelled (~50 out of 400)
    cancelled_indices = set(random.sample(range(400), 50))

    for i in range(400):
        row = i + 2  # data starts at row 2

        order_id = f'ORD-{10001 + i}'
        # Spread dates over about 14 months
        date_val = base_date + timedelta(days=random.randint(0, 425))
        if i in cancelled_indices:
            status = 'Cancelled'
        else:
            status = random.choice(statuses_normal)
        customer_id = f'CUST-{random.randint(1000, 5999):04d}'
        product = random.choice(products)
        category = random.choice(categories)
        units = random.randint(1, 50)
        revenue = round(random.uniform(25.0, 4500.0), 2)
        cost = round(revenue * random.uniform(0.35, 0.85), 2)
        country = random.choice(countries)

        ws.cell(row=row, column=1, value=order_id)
        ws.cell(row=row, column=2, value=date_val)
        ws.cell(row=row, column=3, value=status)
        ws.cell(row=row, column=4, value=customer_id)
        ws.cell(row=row, column=5, value=product)
        ws.cell(row=row, column=6, value=category)
        ws.cell(row=row, column=7, value=units)
        ws.cell(row=row, column=8, value=revenue)
        ws.cell(row=row, column=9, value=cost)
        ws.cell(row=row, column=10, value=country)

    # Format date column
    for r in range(2, 402):
        ws.cell(row=r, column=2).number_format = 'yyyy-mm-dd'

    # Set reasonable column widths
    col_widths = {'A': 12, 'B': 13, 'C': 12, 'D': 14, 'E': 28,
                  'F': 14, 'G': 8, 'H': 12, 'I': 12, 'J': 16}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
