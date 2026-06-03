"""
Reward Script: Track commercial insurance policies and calculate monthly premium allocation
Task ID: calc_fin_insurance_premium_073
Domain: libreoffice_calc
Scoring:
  - Component 1: New column headers G1/H1/I1 present (0.15 pts)
  - Component 2: G2:G20 formulas for Days Coverage (=E-D) (0.20 pts)
  - Component 3: H2:H20 formulas for Monthly Allocation (=F/12) (0.15 pts)
  - Component 4: Data validation dropdown in I2:I20 (0.20 pts)
  - Component 5: Conditional formatting on E2:E20 (0.15 pts)
  - Component 6: Row 1 bold + freeze panes at A2 (0.10 pts)
  - Component 7: Rows sorted by End Date (E col) ascending (0.05 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_insurance_premium_073'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the 'Insurance' sheet exists
    if 'Insurance' not in wb.sheetnames:
        print("CRITICAL: 'Insurance' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Insurance']

    # Component 1: New column headers G1, H1, I1 (0.15 pts)
    # These should only exist in the golden file, not in the initial file
    try:
        g1_val = ws['G1'].value
        h1_val = ws['H1'].value
        i1_val = ws['I1'].value

        headers_found = 0
        if g1_val and 'days' in str(g1_val).lower():
            headers_found += 1
            print(f"PASS: G1 header found: {repr(g1_val)}")
        else:
            print(f"FAIL: G1 expected 'Days Coverage', found: {repr(g1_val)}")

        if h1_val and 'monthly' in str(h1_val).lower():
            headers_found += 1
            print(f"PASS: H1 header found: {repr(h1_val)}")
        else:
            print(f"FAIL: H1 expected 'Monthly Allocation', found: {repr(h1_val)}")

        if i1_val and 'renewal' in str(i1_val).lower():
            headers_found += 1
            print(f"PASS: I1 header found: {repr(i1_val)}")
        else:
            print(f"FAIL: I1 expected 'Renewal Action', found: {repr(i1_val)}")

        if headers_found == 3:
            print("PASS: Component 1 — All 3 new headers present (0.15 pts)")
            total_score += 0.15
        elif headers_found >= 1:
            partial = round(0.05 * headers_found, 2)
            print(f"PARTIAL: Component 1 — {headers_found}/3 headers found ({partial} pts)")
            total_score += partial
        else:
            print("FAIL: Component 1 — No new headers found (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: G2:G20 formulas for Days Coverage (=E-D style) (0.20 pts)
    # Initial file has no formulas in G column
    try:
        formula_count = 0
        valid_formula_count = 0
        for row in range(2, 21):
            cell = ws.cell(row=row, column=7)  # column G
            if cell.value is not None:
                formula_count += 1
                val_str = str(cell.value).upper().replace(' ', '')
                # Accept =E-D or =E/12 or similar formulas involving column E and D
                # Expected: =E2-D2 style
                if '=E' in val_str and '-D' in val_str:
                    valid_formula_count += 1
                elif val_str.startswith('=') and 'E' in val_str and 'D' in val_str:
                    valid_formula_count += 1

        if valid_formula_count >= 15:  # at least 15 of 19 rows
            print(f"PASS: Component 2 — G column formulas (Days Coverage): {valid_formula_count}/19 rows have =E-D formulas (0.20 pts)")
            total_score += 0.20
        elif valid_formula_count >= 8:
            print(f"PARTIAL: Component 2 — G column: {valid_formula_count}/19 rows have =E-D formulas (0.10 pts)")
            total_score += 0.10
        elif formula_count >= 15:
            # Has formulas but wrong type - partial credit
            print(f"PARTIAL: Component 2 — G column has {formula_count} formulas but only {valid_formula_count} are valid =E-D type (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 2 — G column: only {formula_count} cells with values, {valid_formula_count} valid formulas")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: H2:H20 formulas for Monthly Allocation (=F/12 style) (0.15 pts)
    # Initial file has no formulas in H column
    try:
        formula_count = 0
        valid_formula_count = 0
        for row in range(2, 21):
            cell = ws.cell(row=row, column=8)  # column H
            if cell.value is not None:
                formula_count += 1
                val_str = str(cell.value).upper().replace(' ', '')
                # Accept =F/12 or =F2/12 or =F2/(G2/30) style
                if '=F' in val_str and '/12' in val_str:
                    valid_formula_count += 1
                elif '=F' in val_str and '/G' in val_str:  # =F2/(G2/30) style
                    valid_formula_count += 1

        if valid_formula_count >= 15:
            print(f"PASS: Component 3 — H column formulas (Monthly Allocation): {valid_formula_count}/19 rows have valid formulas (0.15 pts)")
            total_score += 0.15
        elif valid_formula_count >= 8:
            print(f"PARTIAL: Component 3 — H column: {valid_formula_count}/19 rows have valid formulas (0.08 pts)")
            total_score += 0.08
        elif formula_count >= 15:
            print(f"PARTIAL: Component 3 — H column has {formula_count} formulas but only {valid_formula_count} are valid type (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 3 — H column: only {formula_count} cells with values, {valid_formula_count} valid formulas")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Data validation dropdown in I2:I20 (0.20 pts)
    # Initial file has no data validations
    try:
        validations = ws.data_validations.dataValidation
        dropdown_found = False
        correct_options = False

        for dv in validations:
            if dv.type == 'list':
                # Check the sqref covers I column
                sqref_str = str(dv.sqref)
                if 'I' in sqref_str:
                    dropdown_found = True
                    # Check required options: Renew, Bid Out, Cancel, Under Review
                    formula_str = str(dv.formula1) if dv.formula1 else ''
                    required_options = ['renew', 'bid out', 'cancel', 'under review']
                    formula_lower = formula_str.lower()
                    options_found = sum(1 for opt in required_options if opt in formula_lower)
                    if options_found >= 3:
                        correct_options = True
                    print(f"  Dropdown found: sqref={sqref_str}, formula1={formula_str}, options_found={options_found}/4")
                    break

        if dropdown_found and correct_options:
            print("PASS: Component 4 — Data validation dropdown in I column with correct options (0.20 pts)")
            total_score += 0.20
        elif dropdown_found:
            print("PARTIAL: Component 4 — Data validation dropdown in I column but missing some options (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 — No list validation found in I column (validations: {len(validations)})")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Conditional formatting on E2:E20 (0.15 pts)
    # Initial file has no conditional formatting
    # Expected: orange for expiring within 90 days, red for expired
    try:
        cf_rules = ws.conditional_formatting
        cf_on_e_col = False
        has_orange_rule = False
        has_red_rule = False

        for cf_range in cf_rules:
            cf_str = str(cf_range)
            if 'E' in cf_str:
                cf_on_e_col = True
                for rule in cf_rules[cf_range]:
                    if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                        try:
                            fill_color = rule.dxf.fill.fgColor.rgb
                            # Orange: FFA500 or similar (FFFFA500)
                            if fill_color and ('FFA500' in fill_color.upper() or 'FF8C00' in fill_color.upper()):
                                has_orange_rule = True
                            # Red: FF0000 or similar (FFFF0000)
                            elif fill_color and ('FF0000' in fill_color.upper()):
                                has_red_rule = True
                        except Exception:
                            pass

        if cf_on_e_col and has_orange_rule and has_red_rule:
            print("PASS: Component 5 — Conditional formatting with orange+red rules on E column (0.15 pts)")
            total_score += 0.15
        elif cf_on_e_col and (has_orange_rule or has_red_rule):
            print(f"PARTIAL: Component 5 — Conditional formatting on E col: orange={has_orange_rule}, red={has_red_rule} (0.08 pts)")
            total_score += 0.08
        elif cf_on_e_col:
            print("PARTIAL: Component 5 — Conditional formatting on E column but color rules not detected (0.05 pts)")
            total_score += 0.05
        else:
            print("FAIL: Component 5 — No conditional formatting on E column")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Row 1 bold headers AND freeze panes at A2 (0.10 pts)
    # Initial file has neither
    try:
        bold_score = 0.0
        freeze_score = 0.0

        # Check row 1 bold
        bold_count = 0
        for col in range(1, 7):  # Check columns A-F (original headers)
            cell = ws.cell(row=1, column=col)
            if cell.font and cell.font.bold:
                bold_count += 1

        if bold_count >= 5:
            bold_score = 0.05
            print(f"PASS: Component 6a — Row 1 bold ({bold_count}/6 header cells bold)")
        else:
            print(f"FAIL: Component 6a — Only {bold_count}/6 header cells are bold")

        # Check freeze panes
        freeze_val = ws.freeze_panes
        if freeze_val and 'A2' in str(freeze_val):
            freeze_score = 0.05
            print(f"PASS: Component 6b — Freeze panes at A2 (value: {freeze_val})")
        else:
            print(f"FAIL: Component 6b — Freeze panes not at A2, found: {freeze_val}")

        comp6_score = bold_score + freeze_score
        if comp6_score > 0:
            print(f"Component 6 score: {comp6_score} pts")
            total_score += comp6_score
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: Rows sorted by End Date (E column) ascending (0.05 pts)
    # Initial file has unsorted rows
    try:
        e_dates = []
        for row in range(2, 21):
            val = ws.cell(row=row, column=5).value
            if val is not None:
                e_dates.append(val)

        if len(e_dates) >= 19:
            is_sorted = all(e_dates[i] <= e_dates[i+1] for i in range(len(e_dates)-1))
            if is_sorted:
                print(f"PASS: Component 7 — Rows sorted by End Date ascending ({len(e_dates)} rows checked) (0.05 pts)")
                total_score += 0.05
            else:
                # Check if at least partially sorted
                sorted_pairs = sum(1 for i in range(len(e_dates)-1) if e_dates[i] <= e_dates[i+1])
                print(f"FAIL: Component 7 — Rows not sorted by End Date ({sorted_pairs}/{len(e_dates)-1} consecutive pairs in order)")
        else:
            print(f"FAIL: Component 7 — Not enough rows with End Date values ({len(e_dates)} found)")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
