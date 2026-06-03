"""
Initial Setup: Insurance policy tracker spreadsheet
Task ID: calc_fin_insurance_premium_073
Domain: libreoffice_calc
"""

import os
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_insurance_premium_073'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Insurance ---
    ws = wb.active
    ws.title = 'Insurance'

    # Headers: Policy# (A), Type (B), Carrier (C), Start Date (D), End Date (E), Annual Premium (F)
    # Columns G, H, I intentionally left empty (task will populate them)
    headers = ['Policy#', 'Type', 'Carrier', 'Start Date', 'End Date', 'Annual Premium']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic insurance policy data (19 rows, sorted roughly by policy type)
    # Mix of past expiries, near-expiring (within 90 days), and future policies
    # Using a reference date of 2026-03-04 (today)
    # Policies: some already expired, some expiring soon, some future
    today = date(2026, 3, 4)

    policies = [
        # Policy#, Type, Carrier, Start Date, End Date, Annual Premium
        ('POL-2023-001', 'General Liability', 'Hartford Insurance',
         date(2025, 4, 1), date(2026, 3, 31), 18500.00),
        ('POL-2023-002', 'Commercial Property', 'Travelers Insurance',
         date(2025, 6, 15), date(2026, 6, 14), 32400.00),
        ('POL-2023-003', 'Workers Compensation', 'Liberty Mutual',
         date(2025, 1, 1), date(2026, 1, 1), 45200.00),
        ('POL-2024-004', 'Commercial Auto', 'Progressive Commercial',
         date(2025, 3, 15), date(2026, 3, 14), 12750.00),
        ('POL-2024-005', 'Umbrella Liability', 'Chubb Group',
         date(2025, 7, 1), date(2026, 6, 30), 8900.00),
        ('POL-2024-006', 'Professional Liability', 'CNA Insurance',
         date(2025, 5, 1), date(2026, 4, 30), 22100.00),
        ('POL-2024-007', 'Directors & Officers', 'AIG Insurance',
         date(2025, 2, 15), date(2026, 2, 14), 55000.00),
        ('POL-2024-008', 'Cyber Liability', 'Coalition Inc',
         date(2025, 8, 1), date(2026, 7, 31), 14200.00),
        ('POL-2024-009', 'Employment Practices', 'Zurich Insurance',
         date(2025, 9, 1), date(2026, 8, 31), 19600.00),
        ('POL-2024-010', 'Product Liability', 'Markel Corporation',
         date(2025, 11, 1), date(2026, 10, 31), 28300.00),
        ('POL-2025-011', 'Commercial Property', 'State Farm Business',
         date(2026, 1, 1), date(2026, 12, 31), 41500.00),
        ('POL-2025-012', 'General Liability', 'Nationwide Business',
         date(2026, 2, 1), date(2027, 1, 31), 16800.00),
        ('POL-2025-013', 'Inland Marine', 'Munich Re Group',
         date(2025, 10, 15), date(2026, 10, 14), 9400.00),
        ('POL-2025-014', 'Business Interruption', 'Swiss Re Group',
         date(2025, 12, 1), date(2026, 11, 30), 33700.00),
        ('POL-2025-015', 'Commercial Auto', 'GEICO Commercial',
         date(2026, 3, 1), date(2027, 2, 28), 11200.00),
        ('POL-2025-016', 'Workers Compensation', 'Berkshire Hathaway',
         date(2026, 4, 1), date(2027, 3, 31), 52800.00),
        ('POL-2025-017', 'Equipment Breakdown', 'FM Global',
         date(2025, 5, 15), date(2026, 5, 14), 7600.00),
        ('POL-2025-018', 'Crime Insurance', 'Hanover Insurance',
         date(2025, 6, 1), date(2026, 5, 31), 5300.00),
        ('POL-2025-019', 'Pollution Liability', 'AIG Environmental',
         date(2026, 1, 15), date(2027, 1, 14), 24900.00),
    ]

    for r, (pol_num, pol_type, carrier, start_dt, end_dt, premium) in enumerate(policies, 2):
        ws.cell(row=r, column=1, value=pol_num)
        ws.cell(row=r, column=2, value=pol_type)
        ws.cell(row=r, column=3, value=carrier)
        ws.cell(row=r, column=4, value=start_dt)
        ws.cell(row=r, column=5, value=end_dt)
        ws.cell(row=r, column=6, value=premium)

    # Format date columns with date number format
    for r in range(2, 21):
        ws.cell(row=r, column=4).number_format = 'yyyy-mm-dd'
        ws.cell(row=r, column=5).number_format = 'yyyy-mm-dd'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 18

    # NOTE: Columns G, H, I are intentionally empty (task will add headers + formulas)
    # NOTE: No bold on row 1 (task will bold headers)
    # NOTE: No freeze panes (task will freeze row 1)
    # NOTE: No conditional formatting (task will add it)
    # NOTE: No data validation (task will add dropdown)
    # NOTE: No currency format on F column (task will add it)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheet: Insurance')
    print('Rows: 19 insurance policies (rows 2-20)')
    print('Columns A-F populated; G, H, I empty')


create_initial()
