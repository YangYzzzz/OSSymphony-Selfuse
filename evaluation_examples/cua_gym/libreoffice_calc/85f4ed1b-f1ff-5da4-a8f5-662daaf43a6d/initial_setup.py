"""
Initial Setup: Define named ranges and VLOOKUP for tax bracket lookup
Task ID: calc_nrv_044
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_044'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Tax Tables ---
    ws_tax = wb.active
    ws_tax.title = 'Tax Tables'

    # Headers
    ws_tax['A1'] = 'Income Threshold'
    ws_tax['B1'] = 'Tax Rate'

    # Style headers
    header_font = Font(bold=True, size=11, name='Calibri')
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    for cell_ref in ['A1', 'B1']:
        cell = ws_tax[cell_ref]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.font = Font(bold=True, size=11, name='Calibri', color="FFFFFF")

    # Tax bracket data (US 2024 approximate brackets)
    thresholds = [0, 10000, 40000, 85000, 165000, 215000]
    rates = [0.10, 0.12, 0.22, 0.24, 0.32, 0.35]

    for i, (threshold, rate) in enumerate(zip(thresholds, rates), start=2):
        ws_tax.cell(row=i, column=1, value=threshold)
        ws_tax.cell(row=i, column=2, value=rate)
        # Format threshold as currency
        ws_tax.cell(row=i, column=1).number_format = '$#,##0'
        # Format rate as percentage
        ws_tax.cell(row=i, column=2).number_format = '0.00%'

    # Set column widths
    ws_tax.column_dimensions['A'].width = 20
    ws_tax.column_dimensions['B'].width = 15

    # --- Sheet 2: Calculator ---
    ws_calc = wb.create_sheet('Calculator')

    ws_calc['C1'] = 'Taxable Income'
    ws_calc['D1'] = 'Marginal Rate'

    # Style headers
    for cell_ref in ['C1', 'D1']:
        cell = ws_calc[cell_ref]
        cell.font = Font(bold=True, size=11, name='Calibri')
        cell.fill = PatternFill(start_color="FF548235", end_color="FF548235", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(bold=True, size=11, name='Calibri', color="FFFFFF")

    # Input value
    ws_calc['C2'] = 92000
    ws_calc['C2'].number_format = '$#,##0'

    # D2 intentionally left empty - task is to add VLOOKUP here
    # Do NOT put any formula or value in D2

    # Add some additional context to make it more realistic
    ws_calc['A1'] = 'Tax Calculator'
    ws_calc['A1'].font = Font(bold=True, size=14, name='Calibri')
    ws_calc['A3'] = 'Instructions:'
    ws_calc['A4'] = 'Enter taxable income in C2 to see the marginal tax rate.'
    ws_calc['A4'].font = Font(italic=True, size=10, name='Calibri', color="808080")

    # Set column widths
    ws_calc.column_dimensions['A'].width = 18
    ws_calc.column_dimensions['B'].width = 5
    ws_calc.column_dimensions['C'].width = 18
    ws_calc.column_dimensions['D'].width = 18

    # NO named ranges defined - that's the task

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
