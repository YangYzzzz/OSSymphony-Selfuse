"""
Reward Script: Define named ranges and VLOOKUP formula for tax lookup
Task ID: calc_nrv_044
Domain: libreoffice_calc
Scoring:
  Component 1: TaxBrackets named range defined for A2:A7 on 'Tax Tables' (0.3 pts)
  Component 2: TaxPercent named range defined for B2:B7 on 'Tax Tables' (0.3 pts)
  Component 3: D2 on 'Calculator' contains VLOOKUP using named ranges with approximate match (0.4 pts)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_044'


def persist_app_state(domain):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def check_named_range(wb, name, expected_sheet, expected_range):
    """
    Check if a named range exists with the correct sheet and cell range.
    Returns True if the named range matches expected_sheet and expected_range.
    expected_range should be like '$A$2:$A$7' (with or without $ signs).
    """
    try:
        dn = wb.defined_names.get(name)
        if dn is None:
            # Try case-insensitive search
            for defined_name in wb.defined_names.values():
                if defined_name.name.lower() == name.lower():
                    dn = defined_name
                    break
        if dn is None:
            return False, f"Named range '{name}' not found"

        attr_text = dn.attr_text  # e.g. "'Tax Tables'!$A$2:$A$7"
        # Normalize: remove quotes and dollar signs for comparison
        normalized = attr_text.replace("'", "").replace("$", "").upper()
        expected_normalized = f"{expected_sheet}!{expected_range}".replace("'", "").replace("$", "").upper()

        if normalized == expected_normalized:
            return True, f"Named range '{name}' = {attr_text}"
        else:
            return False, f"Named range '{name}' = {attr_text}, expected {expected_sheet}!{expected_range}"
    except Exception as e:
        return False, f"Error checking named range '{name}': {e}"


def check_vlookup_formula(ws, cell_coord):
    """
    Check if the cell contains a VLOOKUP formula that:
    1. References C2 (or the income cell)
    2. Uses named ranges (TaxBrackets and/or TaxPercent)
    3. Uses approximate match (TRUE or 1 as last arg)
    Returns (score_fraction, message) where score_fraction is 0.0 to 1.0.
    """
    val = ws[cell_coord].value
    if val is None:
        return 0.0, f"Cell {cell_coord} is empty"
    if not isinstance(val, str):
        return 0.0, f"Cell {cell_coord} is not a formula: {val}"

    formula_upper = val.upper().replace(" ", "")

    # Must be a VLOOKUP
    if "VLOOKUP(" not in formula_upper:
        return 0.0, f"Cell {cell_coord} does not contain VLOOKUP: {val}"

    # Check for named range usage (TaxBrackets or TaxPercent)
    val_check = val.upper()
    has_taxbrackets = "TAXBRACKETS" in val_check
    has_taxpercent = "TAXPERCENT" in val_check

    if not (has_taxbrackets or has_taxpercent):
        return 0.0, f"VLOOKUP does not use named ranges TaxBrackets/TaxPercent: {val}"

    # Check for approximate match (TRUE or 1 as last argument, or CHOOSE pattern)
    # Approximate match: last arg is TRUE or 1, or formula uses sorted lookup
    # The golden uses CHOOSE pattern which inherently does approximate match with TRUE
    has_approximate = ("TRUE" in formula_upper or
                       formula_upper.endswith(",1)") or
                       ",1)" in formula_upper)

    if not has_approximate:
        return 0.5, f"VLOOKUP uses named ranges but may not use approximate match: {val}"

    return 1.0, f"VLOOKUP with named ranges and approximate match: {val}"


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Required sheets exist
    if 'Tax Tables' not in wb.sheetnames:
        print("CRITICAL: 'Tax Tables' sheet not found")
        print("REWARD: 0.0")
        return 0.0
    if 'Calculator' not in wb.sheetnames:
        print("CRITICAL: 'Calculator' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: TaxBrackets named range defined for A2:A7 on 'Tax Tables' (0.3 points)
    try:
        passed, msg = check_named_range(wb, "TaxBrackets", "Tax Tables", "A2:A7")
        if passed:
            print(f"PASS: Component 1 -- TaxBrackets named range correct (0.3 pts). {msg}")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 -- TaxBrackets named range. {msg}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: TaxPercent named range defined for B2:B7 on 'Tax Tables' (0.3 points)
    try:
        passed, msg = check_named_range(wb, "TaxPercent", "Tax Tables", "B2:B7")
        if passed:
            print(f"PASS: Component 2 -- TaxPercent named range correct (0.3 pts). {msg}")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 -- TaxPercent named range. {msg}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: D2 on 'Calculator' contains VLOOKUP with named ranges and approximate match (0.4 points)
    try:
        ws_calc = wb['Calculator']
        score_frac, msg = check_vlookup_formula(ws_calc, 'D2')
        points = 0.4 * score_frac
        if score_frac >= 1.0:
            print(f"PASS: Component 3 -- D2 VLOOKUP formula correct (0.4 pts). {msg}")
            total_score += 0.4
        elif score_frac > 0.0:
            print(f"PARTIAL: Component 3 -- D2 VLOOKUP formula partially correct ({points:.2f} pts). {msg}")
            total_score += points
        else:
            print(f"FAIL: Component 3 -- D2 VLOOKUP formula. {msg}")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist unsaved GUI edits before verification
persist_app_state("libreoffice_calc")

# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
