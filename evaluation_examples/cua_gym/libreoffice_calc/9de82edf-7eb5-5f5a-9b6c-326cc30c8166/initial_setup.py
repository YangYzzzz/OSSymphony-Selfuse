"""
Initial Setup: Goal Seek to find growth rate for 5-year revenue target
Task ID: calc_ggf_043
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_043'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Forecast Sheet ---
    ws = wb.active
    ws.title = 'Forecast'

    # Styling
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    label_font = Font(name='Arial', size=11)
    value_font = Font(name='Arial', size=11)
    title_font = Font(name='Arial', size=14, bold=True, color='2F5496')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    # Title row
    ws.merge_cells('A1:C1')
    ws['A1'] = 'Revenue Growth Forecast Model'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 18

    # Section: Input Parameters
    ws['A2'] = 'INPUT PARAMETERS'
    ws['A2'].font = Font(name='Arial', size=11, bold=True, color='2F5496')

    # Row 2: Base Revenue
    ws['A3'] = 'Base Revenue (Year 0)'
    ws['A3'].font = label_font
    ws['B2'] = 800000
    ws['B2'].number_format = '$#,##0'
    ws['B2'].font = value_font
    ws['B2'].border = thin_border

    # Wait -- let me re-read the task. It says:
    # B2=Base Revenue (800000), B3=Annual Growth Rate (0.10), B8=5-Year Forecast =B2*(1+B3)^5
    # So the layout uses column A for labels and column B for values, with specific rows.
    # Let me match the exact cell references from the task.

    # Clear what we have and redo with exact layout
    ws.unmerge_cells('A1:C1')
    for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=3):
        for cell in row:
            cell.value = None
            cell.font = Font()
            cell.fill = PatternFill()
            cell.alignment = Alignment()
            cell.border = Border()

    # Title
    ws.merge_cells('A1:C1')
    ws['A1'] = 'Revenue Growth Forecast Model'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    # Row 2: Base Revenue
    ws['A2'] = 'Base Revenue (Year 0)'
    ws['A2'].font = label_font
    ws['B2'] = 800000
    ws['B2'].number_format = '$#,##0'
    ws['B2'].font = value_font
    ws['B2'].border = thin_border

    # Row 3: Annual Growth Rate (currently 10%)
    ws['A3'] = 'Annual Growth Rate'
    ws['A3'].font = label_font
    ws['B3'] = 0.10
    ws['B3'].number_format = '0.00%'
    ws['B3'].font = value_font
    ws['B3'].border = thin_border

    # Row 4: blank separator
    ws['A4'] = ''

    # Row 5: Projection header
    ws['A5'] = 'YEAR-BY-YEAR PROJECTIONS'
    ws['A5'].font = Font(name='Arial', size=11, bold=True, color='2F5496')

    # Row 6: Year 1
    ws['A6'] = 'Year 1 Revenue'
    ws['A6'].font = label_font
    ws['B6'] = '=B2*(1+B3)^1'
    ws['B6'].number_format = '$#,##0'
    ws['B6'].border = thin_border

    # Row 7: Year 2
    ws['A7'] = 'Year 2 Revenue'
    ws['A7'].font = label_font
    ws['B7'] = '=B2*(1+B3)^2'
    ws['B7'].number_format = '$#,##0'
    ws['B7'].border = thin_border

    # Row 8: 5-Year Forecast (this is the key cell the task references)
    ws['A8'] = '5-Year Projected Revenue'
    ws['A8'].font = Font(name='Arial', size=11, bold=True)
    ws['B8'] = '=B2*(1+B3)^5'
    ws['B8'].number_format = '$#,##0'
    ws['B8'].font = Font(name='Arial', size=11, bold=True)
    ws['B8'].border = thin_border

    # Additional context rows for realism
    ws['A9'] = ''
    ws['A10'] = 'NOTES'
    ws['A10'].font = Font(name='Arial', size=11, bold=True, color='2F5496')
    ws['A11'] = 'Target: VP of Strategy wants 5-year revenue of $2,000,000'
    ws['A11'].font = Font(name='Arial', size=10, italic=True)
    ws['A12'] = 'Use Tools > Goal Seek to determine required growth rate'
    ws['A12'].font = Font(name='Arial', size=10, italic=True)

    # --- Assumptions Sheet (additional context for realism) ---
    ws2 = wb.create_sheet('Assumptions')
    ws2['A1'] = 'Assumption'
    ws2['B1'] = 'Value'
    ws2['A1'].font = Font(name='Arial', size=11, bold=True)
    ws2['B1'].font = Font(name='Arial', size=11, bold=True)

    assumptions = [
        ['Market growth rate (industry avg)', '8.5%'],
        ['Company historical CAGR', '12.3%'],
        ['Inflation adjustment', '3.2%'],
        ['New product line impact', 'TBD'],
        ['Geographic expansion factor', '1.15x'],
        ['Competitive pressure discount', '-2.0%'],
    ]
    for r, (label, val) in enumerate(assumptions, 2):
        ws2.cell(row=r, column=1, value=label).font = Font(name='Arial', size=10)
        ws2.cell(row=r, column=2, value=val).font = Font(name='Arial', size=10)

    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 15

    # --- Historical Data Sheet ---
    ws3 = wb.create_sheet('Historical')
    ws3['A1'] = 'Year'
    ws3['B1'] = 'Revenue'
    ws3['C1'] = 'Growth %'
    for c in range(1, 4):
        ws3.cell(row=1, column=c).font = Font(name='Arial', size=11, bold=True)

    hist_data = [
        [2020, 520000, None],
        [2021, 565000, '=((B3-B2)/B2)'],
        [2022, 624000, '=((B4-B3)/B3)'],
        [2023, 695000, '=((B5-B4)/B4)'],
        [2024, 762000, '=((B6-B5)/B5)'],
        [2025, 800000, '=((B7-B6)/B6)'],
    ]
    for r, (year, rev, growth) in enumerate(hist_data, 2):
        ws3.cell(row=r, column=1, value=year)
        ws3.cell(row=r, column=2, value=rev)
        ws3['B' + str(r)].number_format = '$#,##0'
        if growth is not None:
            ws3.cell(row=r, column=3, value=growth)
            ws3['C' + str(r)].number_format = '0.00%'

    ws3.column_dimensions['A'].width = 10
    ws3.column_dimensions['B'].width = 18
    ws3.column_dimensions['C'].width = 12

    # Freeze panes on Forecast sheet
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
