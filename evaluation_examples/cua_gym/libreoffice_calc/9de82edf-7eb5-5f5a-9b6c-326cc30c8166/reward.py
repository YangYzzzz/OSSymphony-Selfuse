"""
Reward Script: Goal Seek to find growth rate in B3 that makes B8=2000000
Task ID: calc_ggf_043
Domain: libreoffice_calc
Scoring:
  Component 1 (0.40): B3 value changed to ~0.2012 (Goal Seek result)
  Component 2 (0.25): B3 value yields ~2000000 when used in B2*(1+B3)^5
  Component 3 (0.35): Comment on B3 with Goal Seek result text
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_043'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Forecast sheet must exist
    if 'Forecast' not in wb.sheetnames:
        print("CRITICAL: 'Forecast' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Forecast']

    # Component 1: B3 value changed to ~0.2012 (Goal Seek result) (0.40 points)
    # Initial state has B3=0.10; after Goal Seek it should be ~0.2011-0.2013
    try:
        b3_val = ws['B3'].value
        if b3_val is not None and isinstance(b3_val, (int, float)):
            b3_float = float(b3_val)
            # Goal Seek should find ~0.2012; allow tolerance for solver precision
            # Must NOT be the original 0.10
            if abs(b3_float - 0.10) > 0.01 and abs(b3_float - 0.2012) < 0.02:
                print(f"PASS: Component 1 — B3 value is {b3_float:.6f}, close to expected ~0.2012 (0.40 pts)")
                total_score += 0.40
            else:
                print(f"FAIL: Component 1 — B3 value is {b3_float:.6f}, expected ~0.2012 (not 0.10)")
        else:
            print(f"FAIL: Component 1 — B3 is not a numeric value: {b3_val}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: B3 value yields ~2000000 when used in formula B2*(1+B3)^5 (0.25 points)
    # This validates the Goal Seek math is correct, not just a random value
    try:
        b2_val = ws['B2'].value
        b3_val = ws['B3'].value
        if (b2_val is not None and b3_val is not None
                and isinstance(b2_val, (int, float)) and isinstance(b3_val, (int, float))):
            computed = float(b2_val) * (1 + float(b3_val)) ** 5
            # Allow 1% tolerance for solver precision
            if abs(computed - 2000000) < 20000:
                print(f"PASS: Component 2 — B2*(1+B3)^5 = {computed:.2f}, close to 2000000 (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 2 — B2*(1+B3)^5 = {computed:.2f}, expected ~2000000")
        else:
            print(f"FAIL: Component 2 — B2={b2_val}, B3={b3_val}, need numeric values")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Comment on B3 with Goal Seek result text (0.35 points)
    # The comment should mention "Goal Seek result" and include a percentage
    try:
        b3_comment = ws['B3'].comment
        if b3_comment is not None:
            comment_text = str(b3_comment.text).strip()
            has_goal_seek = 'goal seek' in comment_text.lower()
            has_percentage = any(c == '%' for c in comment_text)
            has_target = '2m' in comment_text.lower() or '2,000,000' in comment_text.lower() or '$2m' in comment_text.lower() or '2000000' in comment_text.lower()

            if has_goal_seek and has_percentage:
                print(f"PASS: Component 3 — B3 has comment: '{comment_text}' (0.35 pts)")
                total_score += 0.35
            else:
                print(f"FAIL: Component 3 — Comment exists but missing required content. "
                      f"goal_seek={has_goal_seek}, percentage={has_percentage}, target={has_target}. "
                      f"Text: '{comment_text}'")
        else:
            print("FAIL: Component 3 — No comment found on B3")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
