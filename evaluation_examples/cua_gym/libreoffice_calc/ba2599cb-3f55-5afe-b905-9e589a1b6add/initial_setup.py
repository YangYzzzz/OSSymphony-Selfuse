"""
Initial Setup: Quarterly KPI data with 5 annual total rows (2019-2023)
Task ID: osworld_calc_annual_pct_change_007
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_annual_pct_change_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- KPI Dashboard sheet ---
    ws = wb.active
    ws.title = "KPI Dashboard"

    # Header row
    headers = ['Year', 'Revenue ($M)', 'Customer Count', 'NPS Score']
    header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2E4057', end_color='FF2E4057', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    ws.row_dimensions[1].height = 28

    # Annual KPI data: 5 years (2019-2023)
    # Revenue in $M, Customer Count, NPS Score (0-100)
    annual_data = [
        # Year, Revenue ($M), Customer Count, NPS Score
        (2019, 48.3,  12450, 42),
        (2020, 51.7,  13820, 38),
        (2021, 61.4,  16340, 51),
        (2022, 74.9,  20110, 58),
        (2023, 89.2,  24680, 63),
    ]

    year_fill = PatternFill(start_color='FFD9E2F3', end_color='FFD9E2F3', fill_type='solid')
    year_font = Font(name='Calibri', bold=True, size=11)
    data_font = Font(name='Calibri', size=11)
    data_align = Alignment(horizontal='center', vertical='center')
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, (year, revenue, customers, nps) in enumerate(annual_data, 2):
        # Year column
        cell_year = ws.cell(row=row_idx, column=1, value=year)
        cell_year.font = year_font
        cell_year.fill = year_fill
        cell_year.alignment = data_align
        cell_year.border = data_border

        # Revenue
        cell_rev = ws.cell(row=row_idx, column=2, value=revenue)
        cell_rev.font = data_font
        cell_rev.alignment = data_align
        cell_rev.border = data_border
        cell_rev.number_format = '#,##0.0'

        # Customer Count
        cell_cust = ws.cell(row=row_idx, column=3, value=customers)
        cell_cust.font = data_font
        cell_cust.alignment = data_align
        cell_cust.border = data_border
        cell_cust.number_format = '#,##0'

        # NPS Score
        cell_nps = ws.cell(row=row_idx, column=4, value=nps)
        cell_nps.font = data_font
        cell_nps.alignment = data_align
        cell_nps.border = data_border
        cell_nps.number_format = '0'

    # Column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    # --- Additional Info sheet ---
    ws2 = wb.create_sheet('Quarterly Detail')

    # Headers for quarterly detail
    q_headers = ['Year', 'Quarter', 'Revenue ($M)', 'Customer Count', 'NPS Score']
    for col, h in enumerate(q_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', bold=True, size=11)
        cell.fill = PatternFill(start_color='FF404040', end_color='FF404040', fill_type='solid')
        cell.font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Quarterly raw data underlying the annual KPIs
    quarterly_data = [
        # Year, Q, Revenue, Customers, NPS
        (2019, 'Q1', 10.8, 2980, 40),
        (2019, 'Q2', 11.5, 3120, 43),
        (2019, 'Q3', 12.6, 3180, 42),
        (2019, 'Q4', 13.4, 3170, 43),
        (2020, 'Q1', 11.9, 3250, 35),
        (2020, 'Q2', 12.4, 3380, 37),
        (2020, 'Q3', 13.5, 3560, 39),
        (2020, 'Q4', 13.9, 3630, 41),
        (2021, 'Q1', 14.2, 3820, 48),
        (2021, 'Q2', 15.1, 4050, 51),
        (2021, 'Q3', 16.0, 4230, 52),
        (2021, 'Q4', 16.1, 4240, 53),
        (2022, 'Q1', 17.3, 4720, 55),
        (2022, 'Q2', 18.5, 5010, 57),
        (2022, 'Q3', 19.8, 5160, 59),
        (2022, 'Q4', 19.3, 5220, 61),
        (2023, 'Q1', 20.4, 5680, 60),
        (2023, 'Q2', 21.9, 6080, 62),
        (2023, 'Q3', 23.1, 6390, 64),
        (2023, 'Q4', 23.8, 6530, 66),
    ]

    for row_idx, (year, qtr, rev, cust, nps) in enumerate(quarterly_data, 2):
        ws2.cell(row=row_idx, column=1, value=year).alignment = Alignment(horizontal='center')
        ws2.cell(row=row_idx, column=2, value=qtr).alignment = Alignment(horizontal='center')
        c_rev = ws2.cell(row=row_idx, column=3, value=rev)
        c_rev.number_format = '#,##0.0'
        c_rev.alignment = Alignment(horizontal='center')
        c_cust = ws2.cell(row=row_idx, column=4, value=cust)
        c_cust.number_format = '#,##0'
        c_cust.alignment = Alignment(horizontal='center')
        ws2.cell(row=row_idx, column=5, value=nps).alignment = Alignment(horizontal='center')

    for col_letter, width in [('A', 10), ('B', 10), ('C', 16), ('D', 16), ('E', 12)]:
        ws2.column_dimensions[col_letter].width = width

    ws2.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
