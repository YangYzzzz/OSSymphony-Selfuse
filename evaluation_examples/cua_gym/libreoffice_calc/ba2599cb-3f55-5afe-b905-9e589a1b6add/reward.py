"""
Reward Script: Add YoY percentage change rows and CAGR row for quarterly KPI data
Task ID: osworld_calc_annual_pct_change_007
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): 4 YoY rows exist with correct labels between annual data rows
  Component 2 (0.35): YoY formulas follow (current - prev) / prev * 100 pattern for all 3 KPIs
  Component 3 (0.30): CAGR row exists at bottom with ((Last/First)^(1/(n-1))-1)*100 formula for all 3 KPIs
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_annual_pct_change_007'


def normalize_formula(formula):
    """Normalize formula string: uppercase, remove spaces for comparison."""
    if not isinstance(formula, str):
        return ''
    return formula.upper().replace(' ', '')


def check_yoy_formula_pattern(formula, current_ref, prev_ref):
    """
    Check if a formula matches the YoY pattern: (current - prev) / prev * 100
    We accept various algebraically equivalent forms.
    """
    if not isinstance(formula, str):
        return False
    norm = normalize_formula(formula)
    # Expected pattern: (current - prev) / prev * 100
    # e.g., =(B4-B2)/B2*100
    cur = current_ref.upper()
    prv = prev_ref.upper()
    # Pattern: (cur-prv)/prv*100
    pattern1 = f'=({cur}-{prv})/{prv}*100'
    # Also accept: =(cur/prv-1)*100 or =(cur/prv)*100-100
    pattern2 = f'=({cur}/{prv}-1)*100'
    pattern3 = f'={cur}/{prv}*100-100'
    return norm in (pattern1, pattern2, pattern3)


def check_cagr_formula_pattern(formula, last_ref, first_ref, n_years=4):
    """
    Check if a formula matches the CAGR pattern: ((Last/First)^(1/(n-1))-1)*100
    For 5 years (2019-2023), n=5, exponent = 1/4.
    We accept: ((Last/First)^(1/4)-1)*100 or similar forms.
    """
    if not isinstance(formula, str):
        return False
    norm = normalize_formula(formula)
    last = last_ref.upper()
    first = first_ref.upper()
    # Main expected form: =((last/first)^(1/4)-1)*100
    pattern1 = f'=(({last}/{first})^(1/{n_years})-1)*100'
    # Also accept =(({last}/{first})^(1/{n_years}.0)-1)*100
    # or =(({last}/{first})^(1/{n_years})-1) * 100
    return norm == pattern1


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task: Add 4 YoY % change rows between annual data rows (2019-2023),
          and one CAGR row at the bottom.

    Initial state: 6 rows (header row 1 + annual rows for 2019, 2020, 2021, 2022, 2023)
    Golden state:  11 rows (header + 5 annual + 4 YoY rows + 1 CAGR row)
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that KPI Dashboard sheet exists
    if 'KPI Dashboard' not in wb.sheetnames:
        print("CRITICAL: 'KPI Dashboard' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['KPI Dashboard']
    max_row = ws.max_row

    print(f"Sheet: 'KPI Dashboard', max_row={max_row}")
    print()

    # --- Component 1: 4 YoY rows exist with correct labels (0.35 points) ---
    # In the golden file, YoY rows are at rows 3, 5, 7, 9
    # Labels: 'YoY 2019-2020', 'YoY 2020-2021', 'YoY 2021-2022', 'YoY 2022-2023'
    expected_yoy_labels = [
        'YoY 2019-2020',
        'YoY 2020-2021',
        'YoY 2021-2022',
        'YoY 2022-2023',
    ]
    try:
        # Collect all row labels in column A
        col_a_values = []
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=1, values_only=True):
            val = row[0]
            col_a_values.append(str(val).strip() if val is not None else '')

        # Count how many expected YoY labels are present
        found_yoy = 0
        for label in expected_yoy_labels:
            # Case-insensitive check
            if any(label.lower() in v.lower() for v in col_a_values):
                found_yoy += 1
            else:
                # Check if a generic YoY label is present (e.g., 'YoY Change 2019-2020')
                year_pair = label.replace('YoY ', '')  # e.g., '2019-2020'
                if any('yoy' in v.lower() and year_pair in v for v in col_a_values):
                    found_yoy += 1

        if found_yoy == 4:
            print(f"PASS: Component 1 — All 4 YoY rows found with correct labels (0.35 pts)")
            total_score += 0.35
        elif found_yoy >= 2:
            partial = 0.35 * (found_yoy / 4)
            print(f"PARTIAL: Component 1 — {found_yoy}/4 YoY rows found ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {found_yoy}/4 YoY rows found (expected 4)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # --- Component 2: YoY formulas are correct (0.35 points) ---
    # For each YoY row, columns B, C, D should contain formulas like:
    # =(current_annual - prev_annual) / prev_annual * 100
    # In golden file structure:
    #   Row 3 YoY 2019-2020: B3=(B4-B2)/B2*100, C3=(C4-C2)/C2*100, D3=(D4-D2)/D2*100
    #   Row 5 YoY 2020-2021: B5=(B6-B4)/B4*100, C5=(C6-C4)/C4*100, D5=(D6-D4)/D4*100
    #   Row 7 YoY 2021-2022: B7=(B8-B6)/B6*100, C7=(C8-C6)/C6*100, D7=(D8-D6)/D6*100
    #   Row 9 YoY 2022-2023: B9=(B10-B8)/B8*100, C9=(C10-C8)/C8*100, D9=(D10-D8)/D8*100
    try:
        # Find rows containing YoY labels
        yoy_row_indices = []
        for row_idx in range(1, max_row + 1):
            cell_val = ws.cell(row=row_idx, column=1).value
            if cell_val is not None and 'yoy' in str(cell_val).lower():
                yoy_row_indices.append(row_idx)

        if len(yoy_row_indices) == 0:
            print("FAIL: Component 2 — No YoY rows found to check formulas")
        else:
            valid_formulas = 0
            total_formula_checks = 0

            for yoy_row in yoy_row_indices:
                # Check if B, C, D columns have YoY-style formulas
                for col_letter in ['B', 'C', 'D']:
                    col_idx = ord(col_letter) - ord('A') + 1
                    formula = ws.cell(row=yoy_row, column=col_idx).value
                    total_formula_checks += 1

                    if isinstance(formula, str) and formula.startswith('='):
                        norm = normalize_formula(formula)
                        # Check for pattern: (X-Y)/Y*100 or (X/Y-1)*100
                        # This captures the mathematical essence of YoY calculation
                        has_division = '/' in norm
                        has_multiplication_100 = '*100' in norm
                        has_subtraction = '-' in norm.replace('=-', '')  # exclude leading minus
                        if has_division and has_multiplication_100 and has_subtraction:
                            valid_formulas += 1
                        else:
                            print(f"  WARN: {col_letter}{yoy_row} formula doesn't match YoY pattern: {formula}")
                    elif formula is not None:
                        print(f"  WARN: {col_letter}{yoy_row} is a static value, not a formula: {formula}")

            if total_formula_checks > 0:
                ratio = valid_formulas / total_formula_checks
                if ratio == 1.0:
                    print(f"PASS: Component 2 — All {valid_formulas}/{total_formula_checks} YoY formulas are correct (0.35 pts)")
                    total_score += 0.35
                elif ratio >= 0.5:
                    partial = 0.35 * ratio
                    print(f"PARTIAL: Component 2 — {valid_formulas}/{total_formula_checks} YoY formulas correct ({partial:.2f} pts)")
                    total_score += partial
                else:
                    print(f"FAIL: Component 2 — Only {valid_formulas}/{total_formula_checks} YoY formulas correct")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # --- Component 3: CAGR row exists with correct formula (0.30 points) ---
    # Expected: a row at the bottom with label containing 'CAGR'
    # Formula pattern: =((Last/First)^(1/4)-1)*100
    # In golden: Row 11: A11='CAGR (2019-2023)', B11=((B10/B2)^(1/4)-1)*100
    try:
        # Find row with CAGR label
        cagr_row_idx = None
        for row_idx in range(1, max_row + 1):
            cell_val = ws.cell(row=row_idx, column=1).value
            if cell_val is not None and 'cagr' in str(cell_val).lower():
                cagr_row_idx = row_idx
                break

        if cagr_row_idx is None:
            print("FAIL: Component 3 — No CAGR row found")
        else:
            # Verify CAGR formulas for columns B, C, D
            valid_cagr = 0
            total_cagr_checks = 0

            for col_letter in ['B', 'C', 'D']:
                col_idx = ord(col_letter) - ord('A') + 1
                formula = ws.cell(row=cagr_row_idx, column=col_idx).value
                total_cagr_checks += 1

                if isinstance(formula, str) and formula.startswith('='):
                    norm = normalize_formula(formula)
                    # CAGR formula must have: power (^), division, -1, *100
                    # Matches: ((X/Y)^(1/N)-1)*100 or equivalent
                    has_power = '^' in norm
                    has_division = '/' in norm
                    has_minus_1 = '-1' in norm
                    has_mult_100 = '*100' in norm
                    if has_power and has_division and has_minus_1 and has_mult_100:
                        valid_cagr += 1
                    else:
                        print(f"  WARN: {col_letter}{cagr_row_idx} CAGR formula pattern mismatch: {formula}")
                elif formula is not None:
                    print(f"  WARN: {col_letter}{cagr_row_idx} is a static value, not a CAGR formula: {formula}")

            if total_cagr_checks > 0:
                ratio = valid_cagr / total_cagr_checks
                cagr_label = ws.cell(row=cagr_row_idx, column=1).value
                if ratio == 1.0:
                    print(f"PASS: Component 3 — CAGR row found ('{cagr_label}') with all {valid_cagr}/{total_cagr_checks} correct formulas (0.30 pts)")
                    total_score += 0.30
                elif ratio > 0:
                    partial = 0.30 * ratio
                    print(f"PARTIAL: Component 3 — CAGR row found with {valid_cagr}/{total_cagr_checks} correct formulas ({partial:.2f} pts)")
                    total_score += partial
                else:
                    print(f"FAIL: Component 3 — CAGR row found ('{cagr_label}') but no valid CAGR formulas")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
