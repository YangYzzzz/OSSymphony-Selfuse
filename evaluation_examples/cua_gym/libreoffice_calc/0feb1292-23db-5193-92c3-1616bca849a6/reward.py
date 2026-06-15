"""
Reward Script: Apply custom four-section number format to cells A2:A5
Task ID: calc_lf_078
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): A2 and A3 have the custom number format applied
  Component 2 (0.3): A4 has the custom number format applied
  Component 3 (0.3): A5 has the custom number format applied
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_078'

# The expected custom number format string
EXPECTED_FORMAT = '[GREEN]#,##0;[RED](#,##0);[BLUE]"-";[BLUE]@'


def persist_app_state(domain: str):
    """Save any unsaved changes in LibreOffice before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the MixedData sheet exists
    if 'MixedData' not in wb.sheetnames:
        print("CRITICAL: Sheet 'MixedData' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['MixedData']

    # Component 1: A2 (positive number) and A3 (negative number) have custom format (0.4 points)
    try:
        a2_fmt = ws['A2'].number_format
        a3_fmt = ws['A3'].number_format
        a2_match = (a2_fmt == EXPECTED_FORMAT)
        a3_match = (a3_fmt == EXPECTED_FORMAT)

        if a2_match and a3_match:
            print(f"PASS: Component 1 — A2 and A3 have custom number format (0.4 pts)")
            print(f"  A2 format: {a2_fmt}")
            print(f"  A3 format: {a3_fmt}")
            total_score += 0.4
        elif a2_match or a3_match:
            print(f"PARTIAL: Component 1 — Only one of A2/A3 has custom format (0.2 pts)")
            print(f"  A2 format: {a2_fmt} (match={a2_match})")
            print(f"  A3 format: {a3_fmt} (match={a3_match})")
            total_score += 0.2
        else:
            print(f"FAIL: Component 1 — A2 and A3 do not have expected format")
            print(f"  A2 format: {a2_fmt}")
            print(f"  A3 format: {a3_fmt}")
            print(f"  Expected:  {EXPECTED_FORMAT}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: A4 (zero value) has custom format (0.3 points)
    try:
        a4_fmt = ws['A4'].number_format
        if a4_fmt == EXPECTED_FORMAT:
            print(f"PASS: Component 2 — A4 has custom number format (0.3 pts)")
            print(f"  A4 format: {a4_fmt}")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — A4 does not have expected format")
            print(f"  A4 format: {a4_fmt}")
            print(f"  Expected:  {EXPECTED_FORMAT}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: A5 (text value) has custom format (0.3 points)
    try:
        a5_fmt = ws['A5'].number_format
        if a5_fmt == EXPECTED_FORMAT:
            print(f"PASS: Component 3 — A5 has custom number format (0.3 pts)")
            print(f"  A5 format: {a5_fmt}")
            total_score += 0.3
        else:
            print(f"FAIL: Component 3 — A5 does not have expected format")
            print(f"  A5 format: {a5_fmt}")
            print(f"  Expected:  {EXPECTED_FORMAT}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist any unsaved state before verification
persist_app_state("libreoffice_calc")

# Run verification
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
