"""
Initial Setup: Create spreadsheet with mixed data types for custom number format task.
Task ID: calc_lf_078
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_078'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: MixedData ---
    ws = wb.active
    ws.title = 'MixedData'

    # Header
    ws.cell(row=1, column=1, value='Value')
    ws['A1'].font = Font(bold=True, size=11, name='Calibri')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Data rows — mixed types for the 4-section format test
    ws.cell(row=2, column=1, value=5000)     # positive number
    ws.cell(row=3, column=1, value=-3200)    # negative number
    ws.cell(row=4, column=1, value=0)        # zero
    ws.cell(row=5, column=1, value='N/A')    # text

    # Set column width for readability
    ws.column_dimensions['A'].width = 18

    # NO custom number format applied — that is the task for the agent

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
