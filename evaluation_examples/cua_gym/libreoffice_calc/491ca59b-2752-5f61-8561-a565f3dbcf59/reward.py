"""
Reward Script: Build a COUNT pivot in Sheet2 showing number of samples per technician per experiment type.
Task ID: osworld_calc_pivot_count_invoice_007
Domain: libreoffice_calc
Scoring:
  Component 1: Sheet2 has a pivot table with data (rows and columns populated)        — 0.3 pts
  Component 2: Correct pivot layout — technicians as rows, experiment types as cols   — 0.4 pts
  Component 3: COUNT values are correct per technician/experiment type pair            — 0.3 pts
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_pivot_count_invoice_007'

# Expected pivot data from Sheet1 ground truth
# 5 technicians x 4 experiment types
EXPECTED_TECHNICIANS = {
    'Dr. Carlos Rivera',
    'Dr. Emma Walsh',
    'Dr. James Okafor',
    'Dr. Priya Nair',
    'Dr. Yuki Tanaka',
}
EXPECTED_EXP_TYPES = {'PCR Analysis', 'Gel Electrophoresis', 'Spectroscopy', 'Cell Culture'}

# Expected counts per technician per experiment type
# Derived from Sheet1 data (30 samples total)
EXPECTED_COUNTS = {
    'Dr. Carlos Rivera': {'PCR Analysis': 1, 'Gel Electrophoresis': 2, 'Spectroscopy': 1, 'Cell Culture': 2},
    'Dr. Emma Walsh':    {'PCR Analysis': 2, 'Gel Electrophoresis': 2, 'Spectroscopy': 1, 'Cell Culture': 1},
    'Dr. James Okafor':  {'PCR Analysis': 1, 'Gel Electrophoresis': 1, 'Spectroscopy': 2, 'Cell Culture': 2},
    'Dr. Priya Nair':    {'PCR Analysis': 2, 'Gel Electrophoresis': 1, 'Spectroscopy': 2, 'Cell Culture': 1},
    'Dr. Yuki Tanaka':   {'PCR Analysis': 2, 'Gel Electrophoresis': 2, 'Spectroscopy': 1, 'Cell Culture': 1},
}
EXPECTED_EXP_TOTALS = {'PCR Analysis': 8, 'Gel Electrophoresis': 8, 'Spectroscopy': 7, 'Cell Culture': 7}
EXPECTED_GRAND_TOTAL = 30


def verify_task(file_path):
    """
    Verify that Sheet2 contains a COUNT pivot table with:
      - Technicians as rows
      - Experiment Types as columns
      - Correct count values
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify Sheet2 exists
    if 'Sheet2' not in wb.sheetnames and len(wb.sheetnames) < 2:
        print("FAIL: Sheet2 does not exist in the workbook")
        print("REWARD: 0.0")
        return 0.0

    # Get Sheet2 (by name or by index)
    if 'Sheet2' in wb.sheetnames:
        ws2 = wb['Sheet2']
    else:
        ws2 = wb.worksheets[1]

    # -----------------------------------------------------------------------
    # Component 1: Sheet2 has a pivot table with data (0.3 points)
    # Check that Sheet2 has meaningful tabular data with multiple rows/columns.
    # In the initial state, Sheet2 is empty (A1:A1 with no data).
    # -----------------------------------------------------------------------
    try:
        # Gather all non-None values in Sheet2
        non_empty_cells = []
        for row in ws2.iter_rows():
            for cell in row:
                if cell.value is not None:
                    non_empty_cells.append((cell.row, cell.column, cell.value))

        # Require at least 3 rows with data (header row + at least 2 data rows)
        # and at least 3 columns (technician col + at least 2 experiment type cols)
        rows_with_data = set(r for r, c, v in non_empty_cells)
        cols_with_data = set(c for r, c, v in non_empty_cells)

        if len(rows_with_data) >= 3 and len(cols_with_data) >= 3:
            print(f"PASS: Component 1 — Sheet2 has pivot data: {len(rows_with_data)} rows, {len(cols_with_data)} cols with data (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — Sheet2 has insufficient data: {len(rows_with_data)} rows, {len(cols_with_data)} cols with data")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Correct pivot layout — technicians as rows,
    #              experiment types as columns (0.4 points)
    # We look for a header row that contains at least 3 of the 4 experiment
    # type names, and a column whose cells contain at least 3 of the 5
    # technician names.
    # -----------------------------------------------------------------------
    try:
        # Find the header row containing experiment type names
        header_row_idx = None
        exp_col_map = {}  # exp_type -> column index

        for row in ws2.iter_rows():
            row_values = [str(cell.value).strip() if cell.value is not None else '' for cell in row]
            found_exp_types = set(v for v in row_values if v in EXPECTED_EXP_TYPES)
            if len(found_exp_types) >= 3:
                header_row_idx = row[0].row
                # Build col map
                for cell in row:
                    val = str(cell.value).strip() if cell.value is not None else ''
                    if val in EXPECTED_EXP_TYPES:
                        exp_col_map[val] = cell.column
                break

        # Find the technician column — search for cells containing technician names
        tech_col_idx = None
        tech_row_map = {}  # technician_name -> row index

        if header_row_idx is not None:
            # Search column 1 first, then any column
            for col_idx in range(1, ws2.max_column + 1):
                found_techs = []
                for row_idx in range(header_row_idx + 1, ws2.max_row + 1):
                    cell_val = ws2.cell(row=row_idx, column=col_idx).value
                    if cell_val is not None and str(cell_val).strip() in EXPECTED_TECHNICIANS:
                        found_techs.append((row_idx, str(cell_val).strip()))
                if len(found_techs) >= 3:
                    tech_col_idx = col_idx
                    for row_idx, tech_name in found_techs:
                        tech_row_map[tech_name] = row_idx
                    break

        if header_row_idx is not None and tech_col_idx is not None and len(exp_col_map) >= 3 and len(tech_row_map) >= 3:
            print(f"PASS: Component 2 — Correct layout: header row {header_row_idx}, "
                  f"found {len(exp_col_map)} exp type cols, {len(tech_row_map)} technician rows (0.4 pts)")
            total_score += 0.4
        else:
            missing_parts = []
            if header_row_idx is None:
                missing_parts.append("no header row with experiment types found")
            elif len(exp_col_map) < 3:
                missing_parts.append(f"only {len(exp_col_map)} experiment type columns found (need >= 3)")
            if tech_col_idx is None:
                missing_parts.append("no technician column found")
            elif len(tech_row_map) < 3:
                missing_parts.append(f"only {len(tech_row_map)} technician rows found (need >= 3)")
            print(f"FAIL: Component 2 — Layout check failed: {'; '.join(missing_parts)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")
        header_row_idx = None
        tech_col_idx = None
        exp_col_map = {}
        tech_row_map = {}

    # -----------------------------------------------------------------------
    # Component 3: COUNT values are correct per technician/experiment type (0.3 pts)
    # Verify that the pivot cell values match the expected counts derived from
    # Sheet1. We require at least 80% of the expected cross-cells to be correct.
    # -----------------------------------------------------------------------
    try:
        if not (header_row_idx and tech_col_idx and exp_col_map and tech_row_map):
            print("FAIL: Component 3 — Skipped because layout (Component 2) was not verified")
        else:
            correct_cells = 0
            total_cells = 0
            mismatches = []

            for tech_name, expected_row_idx in tech_row_map.items():
                for exp_type, expected_col_idx in exp_col_map.items():
                    expected_count = EXPECTED_COUNTS.get(tech_name, {}).get(exp_type)
                    if expected_count is None:
                        continue
                    actual_val = ws2.cell(row=expected_row_idx, column=expected_col_idx).value
                    total_cells += 1
                    if actual_val is not None and int(actual_val) == expected_count:
                        correct_cells += 1
                    else:
                        mismatches.append(
                            f"{tech_name}/{exp_type}: expected {expected_count}, got {actual_val}"
                        )

            if total_cells == 0:
                print("FAIL: Component 3 — No cells to verify (layout mismatch)")
            else:
                accuracy = correct_cells / total_cells
                if accuracy >= 0.8:
                    print(f"PASS: Component 3 — COUNT values correct: {correct_cells}/{total_cells} cells match (0.3 pts)")
                    total_score += 0.3
                else:
                    print(f"FAIL: Component 3 — COUNT values incorrect: only {correct_cells}/{total_cells} cells match")
                    if mismatches:
                        print(f"  Mismatches: {mismatches[:5]}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
