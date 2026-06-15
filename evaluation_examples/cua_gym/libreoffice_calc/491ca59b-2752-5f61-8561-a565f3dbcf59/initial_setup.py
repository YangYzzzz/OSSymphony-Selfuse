"""
Initial Setup: Lab sample tracking spreadsheet for pivot table creation
Task ID: osworld_calc_pivot_count_invoice_007
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_pivot_count_invoice_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Lab Data ---
    ws1 = wb.active
    ws1.title = 'Sheet1'

    # Headers
    headers = ['Sample ID', 'Technician', 'Experiment Type', 'Date', 'Result Status', 'Processing Time (hrs)']
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws1.column_dimensions['A'].width = 14
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 22
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 16
    ws1.column_dimensions['F'].width = 22

    # Realistic lab data: 5 technicians x 4 experiment types = varied distribution
    # Technicians: Dr. Emma Walsh, Dr. Carlos Rivera, Dr. Priya Nair, Dr. James Okafor, Dr. Yuki Tanaka
    # Experiment Types: PCR Analysis, Gel Electrophoresis, Spectroscopy, Cell Culture
    data = [
        # Sample ID, Technician, Experiment Type, Date, Result Status, Processing Time
        ['SMP-001', 'Dr. Emma Walsh',     'PCR Analysis',       '2025-01-07', 'Completed', 2.5],
        ['SMP-002', 'Dr. Carlos Rivera',  'Gel Electrophoresis','2025-01-08', 'Completed', 1.8],
        ['SMP-003', 'Dr. Priya Nair',     'Spectroscopy',       '2025-01-09', 'Completed', 3.2],
        ['SMP-004', 'Dr. James Okafor',   'Cell Culture',       '2025-01-10', 'Pending',   4.0],
        ['SMP-005', 'Dr. Yuki Tanaka',    'PCR Analysis',       '2025-01-11', 'Completed', 2.1],
        ['SMP-006', 'Dr. Emma Walsh',     'Gel Electrophoresis','2025-01-13', 'Completed', 1.5],
        ['SMP-007', 'Dr. Carlos Rivera',  'Spectroscopy',       '2025-01-14', 'Failed',    3.8],
        ['SMP-008', 'Dr. Priya Nair',     'Cell Culture',       '2025-01-15', 'Completed', 5.0],
        ['SMP-009', 'Dr. James Okafor',   'PCR Analysis',       '2025-01-16', 'Completed', 2.3],
        ['SMP-010', 'Dr. Yuki Tanaka',    'Gel Electrophoresis','2025-01-17', 'Completed', 1.9],
        ['SMP-011', 'Dr. Emma Walsh',     'Spectroscopy',       '2025-01-20', 'Completed', 3.5],
        ['SMP-012', 'Dr. Carlos Rivera',  'Cell Culture',       '2025-01-21', 'Pending',   4.7],
        ['SMP-013', 'Dr. Priya Nair',     'PCR Analysis',       '2025-01-22', 'Completed', 2.0],
        ['SMP-014', 'Dr. James Okafor',   'Gel Electrophoresis','2025-01-23', 'Completed', 1.6],
        ['SMP-015', 'Dr. Yuki Tanaka',    'Spectroscopy',       '2025-01-24', 'Failed',    4.1],
        ['SMP-016', 'Dr. Emma Walsh',     'Cell Culture',       '2025-01-27', 'Completed', 5.5],
        ['SMP-017', 'Dr. Carlos Rivera',  'PCR Analysis',       '2025-01-28', 'Completed', 2.8],
        ['SMP-018', 'Dr. Priya Nair',     'Gel Electrophoresis','2025-01-29', 'Completed', 1.4],
        ['SMP-019', 'Dr. James Okafor',   'Spectroscopy',       '2025-01-30', 'Completed', 3.0],
        ['SMP-020', 'Dr. Yuki Tanaka',    'Cell Culture',       '2025-01-31', 'Completed', 4.9],
        ['SMP-021', 'Dr. Emma Walsh',     'PCR Analysis',       '2025-02-03', 'Completed', 2.2],
        ['SMP-022', 'Dr. Carlos Rivera',  'Gel Electrophoresis','2025-02-04', 'Completed', 1.7],
        ['SMP-023', 'Dr. Priya Nair',     'Spectroscopy',       '2025-02-05', 'Failed',    3.6],
        ['SMP-024', 'Dr. James Okafor',   'Cell Culture',       '2025-02-06', 'Completed', 4.3],
        ['SMP-025', 'Dr. Yuki Tanaka',    'PCR Analysis',       '2025-02-07', 'Completed', 2.4],
        ['SMP-026', 'Dr. Emma Walsh',     'Gel Electrophoresis','2025-02-10', 'Completed', 1.3],
        ['SMP-027', 'Dr. Carlos Rivera',  'Cell Culture',       '2025-02-11', 'Completed', 4.5],
        ['SMP-028', 'Dr. Priya Nair',     'PCR Analysis',       '2025-02-12', 'Completed', 2.6],
        ['SMP-029', 'Dr. James Okafor',   'Spectroscopy',       '2025-02-13', 'Completed', 3.4],
        ['SMP-030', 'Dr. Yuki Tanaka',    'Gel Electrophoresis','2025-02-14', 'Completed', 1.6],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Freeze header row
    ws1.freeze_panes = 'A2'

    # --- Sheet 2: Empty (agent will build pivot here) ---
    ws2 = wb.create_sheet('Sheet2')
    # Sheet2 is intentionally left empty — the agent must build the pivot table here

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
