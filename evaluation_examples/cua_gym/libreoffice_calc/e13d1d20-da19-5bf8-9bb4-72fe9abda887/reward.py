"""
Reward Script: Use VLOOKUP to add supplier country to each inventory row,
               and create a pivot table in Sheet2 summarizing total inventory
               value by supplier country.
Task ID: osworld_calc_vlookup_pivot_combined_005
Domain: libreoffice_calc
Scoring:
  - Component 1: Column F header is 'Country' in Sheet1 (0.10 pts)
  - Component 2: VLOOKUP formulas present in Sheet1 column F rows 2-21 (0.50 pts)
  - Component 3: Sheet2 exists with pivot table header row (0.10 pts)
  - Component 4: Sheet2 has all 6 supplier countries with correct total inventory values (0.30 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_vlookup_pivot_combined_005'

# Expected pivot table data in Sheet2 (Supplier Country -> Total Inventory Value)
# Computed from: Stock Qty * Unit Cost grouped by Supplier Country
EXPECTED_PIVOT = {
    'China': 39235,
    'Germany': 29962.5,
    'Japan': 29540.5,
    'South Korea': 89597.15,
    'Taiwan': 48647.7,
    'United States': 61307.9,
}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sheet1 must exist
    if 'Sheet1' not in wb.sheetnames:
        print("CRITICAL: Sheet1 not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws1 = wb['Sheet1']

    # Component 1: Column F header is 'Country' (0.10 points)
    # Initial: F1 is None. Golden: F1 is 'Country'.
    try:
        f1_value = ws1['F1'].value
        if f1_value is not None and str(f1_value).strip().lower() == 'country':
            print(f"PASS: Component 1 — Column F header is 'Country' (found: '{f1_value}') (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 — Column F header expected 'Country', found: '{f1_value}'")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: VLOOKUP formulas in Sheet1 column F rows 2-21 (0.50 points)
    # Initial: column F rows 2-21 are all None.
    # Golden: each cell contains a VLOOKUP formula referencing column B and $G:$H.
    # We require at least 15 of the 20 rows to have VLOOKUP formulas for partial credit,
    # and award full credit when all 20 rows have VLOOKUP formulas.
    try:
        data_row_count = 0  # rows 2-21 = 20 data rows
        vlookup_count = 0
        for row in range(2, 22):
            cell_val = ws1.cell(row=row, column=6).value
            data_row_count += 1
            if cell_val is not None and isinstance(cell_val, str):
                val_upper = cell_val.upper().replace(' ', '')
                if 'VLOOKUP' in val_upper:
                    vlookup_count += 1

        if vlookup_count == 20:
            print(f"PASS: Component 2 — All 20 rows (F2:F21) have VLOOKUP formulas (0.50 pts)")
            total_score += 0.50
        elif vlookup_count >= 15:
            partial = round(0.50 * (vlookup_count / 20), 2)
            print(f"PARTIAL: Component 2 — {vlookup_count}/20 rows have VLOOKUP formulas ({partial} pts)")
            total_score += partial
        elif vlookup_count > 0:
            partial = round(0.50 * (vlookup_count / 20), 2)
            print(f"PARTIAL: Component 2 — {vlookup_count}/20 rows have VLOOKUP formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No VLOOKUP formulas found in column F rows 2-21")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Sheet2 exists with pivot table header row (0.10 points)
    # Initial: Sheet2 does not exist.
    # Golden: Sheet2 exists with headers 'Supplier Country' and 'Total Inventory Value' in row 1.
    try:
        if 'Sheet2' not in wb.sheetnames:
            print("FAIL: Component 3 — Sheet2 does not exist")
        else:
            ws2 = wb['Sheet2']
            h1 = ws2.cell(row=1, column=1).value
            h2 = ws2.cell(row=1, column=2).value
            if (h1 is not None and 'country' in str(h1).lower()) and \
               (h2 is not None and ('inventory' in str(h2).lower() or 'value' in str(h2).lower())):
                print(f"PASS: Component 3 — Sheet2 exists with headers '{h1}', '{h2}' (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 3 — Sheet2 has unexpected headers: '{h1}', '{h2}'")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Sheet2 has all 6 supplier countries with correct total inventory values (0.30 points)
    # Initial: Sheet2 does not exist (all checks fail).
    # Golden: Sheet2 has 6 rows of country-value pairs matching EXPECTED_PIVOT.
    try:
        if 'Sheet2' not in wb.sheetnames:
            print("FAIL: Component 4 — Sheet2 does not exist")
        else:
            ws2 = wb['Sheet2']
            # Read the pivot table data from Sheet2 (rows 2 onwards)
            found_pivot = {}
            for row in range(2, ws2.max_row + 1):
                country_cell = ws2.cell(row=row, column=1).value
                value_cell = ws2.cell(row=row, column=2).value
                if country_cell is not None:
                    found_pivot[str(country_cell).strip()] = value_cell

            correct_count = 0
            tolerance = 1.0  # allow $1 tolerance for rounding

            for country, expected_val in EXPECTED_PIVOT.items():
                if country in found_pivot:
                    actual_val = found_pivot[country]
                    if actual_val is not None:
                        try:
                            if abs(float(actual_val) - float(expected_val)) <= tolerance:
                                correct_count += 1
                            else:
                                print(f"  MISMATCH: {country}: expected {expected_val}, got {actual_val}")
                        except (ValueError, TypeError):
                            print(f"  TYPE_ERROR: {country}: got non-numeric value '{actual_val}'")
                    else:
                        print(f"  MISSING_VALUE: {country}: value is None")
                else:
                    print(f"  MISSING_COUNTRY: '{country}' not found in Sheet2")

            if correct_count == 6:
                print(f"PASS: Component 4 — All 6 supplier countries with correct inventory values (0.30 pts)")
                total_score += 0.30
            elif correct_count > 0:
                partial = round(0.30 * (correct_count / 6), 2)
                print(f"PARTIAL: Component 4 — {correct_count}/6 countries correct ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 4 — No correct country-value pairs found in Sheet2")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point: test against canonical artifact path in the VM environment
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
