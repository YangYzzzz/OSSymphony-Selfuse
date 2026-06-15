"""
Initial Setup: VLOOKUP + Pivot Table task — Inventory with Supplier Code/Country
Task ID: osworld_calc_vlookup_pivot_combined_005
Domain: libreoffice_calc

Creates Sheet1 with inventory data (Item ID, Supplier Code, Product, Stock Qty,
Unit Cost) and a Supplier Code→Country lookup table in columns G-H.
Sheet2 is intentionally absent (agent must create it).
The Country column (F) is intentionally absent (agent must add it via VLOOKUP).
"""

import os
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_vlookup_pivot_combined_005'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ── Sheet1: Inventory ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Sheet1"

    # --- Headers (A1:E1) for inventory data ---
    inv_headers = ["Item ID", "Supplier Code", "Product", "Stock Qty", "Unit Cost"]
    for col, h in enumerate(inv_headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # --- Inventory data rows (realistic content) ---
    # NOTE: Country column (F) is intentionally absent — agent must add via VLOOKUP
    inventory_data = [
        ["ITM-001", "SUP-CN",  "Wireless Keyboard",       320,  28.50],
        ["ITM-002", "SUP-DE",  "Ergonomic Mouse",          185,  45.00],
        ["ITM-003", "SUP-US",  "USB-C Hub 7-port",         210,  34.99],
        ["ITM-004", "SUP-JP",  "Mechanical Keyboard",       95, 129.00],
        ["ITM-005", "SUP-CN",  "27\" Monitor Stand",       140,  62.75],
        ["ITM-006", "SUP-TW",  "PCIe SSD 1TB",             230,  89.99],
        ["ITM-007", "SUP-DE",  "Laptop Cooling Pad",       175,  38.50],
        ["ITM-008", "SUP-US",  "Webcam 1080p",             310,  55.00],
        ["ITM-009", "SUP-JP",  "HDMI 2.1 Cable 3m",        450,  12.99],
        ["ITM-010", "SUP-TW",  "DDR5 RAM 16GB",            165, 110.00],
        ["ITM-011", "SUP-CN",  "Network Switch 8-port",    120,  74.50],
        ["ITM-012", "SUP-KR",  "OLED Monitor 34\"",         55, 680.00],
        ["ITM-013", "SUP-US",  "Noise-Cancelling Headset", 200,  95.00],
        ["ITM-014", "SUP-KR",  "Portable SSD 2TB",         285,  79.99],
        ["ITM-015", "SUP-DE",  "USB 3.2 Docking Station",  100, 149.00],
        ["ITM-016", "SUP-TW",  "Mini PC Barebones",         40, 245.00],
        ["ITM-017", "SUP-JP",  "Thunderbolt 4 Hub",        130,  88.00],
        ["ITM-018", "SUP-CN",  "IP Security Camera",       295,  42.00],
        ["ITM-019", "SUP-KR",  "Curved Gaming Monitor 27\"", 70, 420.00],
        ["ITM-020", "SUP-US",  "VoIP Conference Speaker",   90, 199.00],
    ]

    for r, row_data in enumerate(inventory_data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # --- Lookup table in columns G-H: Supplier Code → Country ---
    lookup_headers = ["Supplier Code", "Country"]
    for col, h in enumerate(lookup_headers, 7):  # columns G=7, H=8
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    lookup_data = [
        ["SUP-CN", "China"],
        ["SUP-DE", "Germany"],
        ["SUP-JP", "Japan"],
        ["SUP-KR", "South Korea"],
        ["SUP-TW", "Taiwan"],
        ["SUP-US", "United States"],
    ]
    for r, row_data in enumerate(lookup_data, 2):
        for c, val in enumerate(row_data, 7):  # G=7, H=8
            ws1.cell(row=r, column=c, value=val)

    # Adjust column widths for readability
    ws1.column_dimensions["A"].width = 10
    ws1.column_dimensions["B"].width = 14
    ws1.column_dimensions["C"].width = 28
    ws1.column_dimensions["D"].width = 11
    ws1.column_dimensions["E"].width = 11
    ws1.column_dimensions["G"].width = 14
    ws1.column_dimensions["H"].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
