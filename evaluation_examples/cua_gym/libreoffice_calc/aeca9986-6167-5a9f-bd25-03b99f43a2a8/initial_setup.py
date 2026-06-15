"""
Initial Setup: Sales tracker with monthly targets and actuals data.
Task ID: calc_nrv_021
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_021'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Tracker"

    # --- Headers ---
    headers = ['Month', 'Region', 'Product', 'Sales Rep', 'Target', 'Actual', 'Variance', 'Total Variance']
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Monthly Data (rows 2-12, 11 months Jan-Nov 2025) ---
    months = [
        'January', 'February', 'March', 'April', 'May', 'June',
        'July', 'August', 'September', 'October', 'November'
    ]
    regions = [
        'Northeast', 'Southeast', 'Midwest', 'West Coast', 'Northeast',
        'Southeast', 'Midwest', 'West Coast', 'Northeast', 'Southeast', 'Midwest'
    ]
    products = [
        'Widget Pro', 'Widget Lite', 'Widget Pro', 'Widget Max', 'Widget Lite',
        'Widget Max', 'Widget Pro', 'Widget Lite', 'Widget Max', 'Widget Pro', 'Widget Lite'
    ]
    reps = [
        'Sarah Chen', 'Marcus Johnson', 'Priya Patel', 'David Kim', 'Elena Rodriguez',
        'James Wilson', 'Aisha Mohammed', 'Robert Taylor', 'Lisa Wang', 'Carlos Mendez', 'Hannah Scott'
    ]
    targets = [45000, 38000, 52000, 41000, 47000, 35000, 55000, 39000, 48000, 42000, 50000]
    actuals = [47250, 36100, 54800, 39500, 49200, 37800, 52100, 41300, 50750, 40200, 53400]

    for i, month in enumerate(months):
        row = i + 2
        ws.cell(row=row, column=1, value=month)
        ws.cell(row=row, column=2, value=regions[i])
        ws.cell(row=row, column=3, value=products[i])
        ws.cell(row=row, column=4, value=reps[i])
        target_cell = ws.cell(row=row, column=5, value=targets[i])
        target_cell.number_format = '$#,##0'
        actual_cell = ws.cell(row=row, column=6, value=actuals[i])
        actual_cell.number_format = '$#,##0'

    # G2 and H2 intentionally left EMPTY (task requires agent to fill these)

    # --- Column widths ---
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 16

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
