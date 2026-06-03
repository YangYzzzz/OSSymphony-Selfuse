"""
Reward Script: Sales tracker named ranges and variance formulas
Task ID: calc_nrv_021
Domain: libreoffice_calc
Scoring:
  Component 1: Named range 'Targets' refers to $E$2:$E$12 (0.25 pts)
  Component 2: Named range 'Actuals' refers to $F$2:$F$12 (0.25 pts)
  Component 3: G2 contains variance formula for first month (0.25 pts)
  Component 4: H2 contains SUMPRODUCT total variance formula using named ranges (0.25 pts)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_021'


def normalize_ref(ref_str):
    """Normalize a defined name reference for comparison.
    Strips quotes around sheet names and normalizes to uppercase.
    E.g. \"'Sales Tracker'!$E$2:$E$12\" -> \"SALES TRACKER!$E$2:$E$12\"
    """
    s = ref_str.strip().upper()
    # Remove surrounding quotes on the sheet name portion
    s = re.sub(r"'([^']+)'!", r"\1!", s)
    return s


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Component 1: Named range 'Targets' refers to $E$2:$E$12 (0.25 points)
    try:
        found_targets = False
        for name, dn in wb.defined_names.items():
            if name.lower() == 'targets':
                ref = normalize_ref(dn.attr_text)
                # Accept any sheet name as long as the cell range is $E$2:$E$12
                if '$E$2:$E$12' in ref:
                    found_targets = True
                    print(f"PASS: Component 1 -- Named range 'Targets' found: {dn.attr_text} (0.25 pts)")
                    total_score += 0.25
                else:
                    print(f"FAIL: Component 1 -- Named range 'Targets' has wrong range: {dn.attr_text}, expected $E$2:$E$12")
                break
        if not found_targets and total_score < 0.25:
            print("FAIL: Component 1 -- Named range 'Targets' not found")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Named range 'Actuals' refers to $F$2:$F$12 (0.25 points)
    try:
        found_actuals = False
        for name, dn in wb.defined_names.items():
            if name.lower() == 'actuals':
                ref = normalize_ref(dn.attr_text)
                if '$F$2:$F$12' in ref:
                    found_actuals = True
                    print(f"PASS: Component 2 -- Named range 'Actuals' found: {dn.attr_text} (0.25 pts)")
                    total_score += 0.25
                else:
                    print(f"FAIL: Component 2 -- Named range 'Actuals' has wrong range: {dn.attr_text}, expected $F$2:$F$12")
                break
        if not found_actuals and total_score < 0.5:
            print("FAIL: Component 2 -- Named range 'Actuals' not found")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: G2 contains a variance formula (Actual - Target) for first month (0.25 points)
    try:
        g2_val = ws['G2'].value
        if g2_val is not None and isinstance(g2_val, str) and g2_val.startswith('='):
            formula_upper = g2_val.upper().replace(' ', '')
            # The formula should compute Actual minus Target for the first month.
            # Acceptable forms: =F2-E2, =Actuals-Targets (if single cell ref via named range)
            # Also accept INDEX-based: =INDEX(Actuals,1)-INDEX(Targets,1)
            is_variance = False
            # Direct cell reference: =F2-E2
            if 'F2' in formula_upper and 'E2' in formula_upper and '-' in formula_upper:
                is_variance = True
            # Named range single-cell access patterns
            if 'ACTUALS' in formula_upper and 'TARGETS' in formula_upper and '-' in formula_upper:
                is_variance = True
            # INDEX-based
            if 'INDEX' in formula_upper and '-' in formula_upper:
                is_variance = True

            if is_variance:
                print(f"PASS: Component 3 -- G2 has variance formula: {g2_val} (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 3 -- G2 formula does not compute variance (Actual-Target): {g2_val}")
        else:
            print(f"FAIL: Component 3 -- G2 is not a formula, found: {g2_val}")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: H2 contains SUMPRODUCT formula for total variance using named ranges (0.25 points)
    try:
        h2_val = ws['H2'].value
        if h2_val is not None and isinstance(h2_val, str) and h2_val.startswith('='):
            formula_upper = h2_val.upper().replace(' ', '')
            # Expected: =SUMPRODUCT(Actuals-Targets) or similar
            # Must use SUMPRODUCT and reference named ranges (or equivalent range)
            has_sumproduct = 'SUMPRODUCT' in formula_upper
            # Accept named ranges or direct cell ranges for the subtraction
            uses_ranges = (
                ('ACTUALS' in formula_upper and 'TARGETS' in formula_upper) or
                ('F2:F12' in formula_upper and 'E2:E12' in formula_upper) or
                ('$F$2:$F$12' in formula_upper and '$E$2:$E$12' in formula_upper)
            )
            # Also accept SUM-based total variance formulas as valid alternatives
            has_sum = 'SUM(' in formula_upper and '-' in formula_upper

            if has_sumproduct and uses_ranges:
                print(f"PASS: Component 4 -- H2 has SUMPRODUCT total variance formula: {h2_val} (0.25 pts)")
                total_score += 0.25
            elif has_sumproduct:
                # SUMPRODUCT present but not using named ranges directly -- partial acceptance
                # Check if it at least computes a difference
                if '-' in formula_upper:
                    print(f"PASS: Component 4 -- H2 has SUMPRODUCT difference formula: {h2_val} (0.25 pts)")
                    total_score += 0.25
                else:
                    print(f"FAIL: Component 4 -- H2 has SUMPRODUCT but no subtraction: {h2_val}")
            elif has_sum and uses_ranges:
                # SUM-based alternative for total variance
                print(f"PASS: Component 4 -- H2 has SUM-based total variance formula: {h2_val} (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 4 -- H2 does not have expected total variance formula: {h2_val}")
        else:
            print(f"FAIL: Component 4 -- H2 is not a formula, found: {h2_val}")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
