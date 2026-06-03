"""
Reward Script: Calculate percentile rank in F2 using RANK and COUNT
Task ID: calc_fmb_percentile_benchmarking_076
Domain: libreoffice_calc

Scoring Rubric:
  Component 1 (0.5): F2 contains a formula referencing RANK and COUNT over E2:E201
  Component 2 (0.3): F2 formula uses the correct 1-RANK/COUNT structure with correct range and ordering
  Component 3 (0.2): No other cells were modified (only F2 changed from initial state)

Task: Put =1-RANK(E2,E$2:E$201,0)/COUNT(E$2:E$201) in cell F2 of 'Performance Review' sheet.
"""

import os
import openpyxl
import re

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmb_percentile_benchmarking_076'


def normalize_formula(formula):
    """Normalize formula for comparison: uppercase, remove spaces."""
    if not isinstance(formula, str):
        return ''
    return formula.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify 'Performance Review' sheet exists
    if 'Performance Review' not in wb.sheetnames:
        print("FAIL: Sheet 'Performance Review' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Performance Review']

    # Component 1: F2 contains a formula that uses both RANK and COUNT
    # referencing column E rows 2-201 (0.5 points)
    # This FAILS on initial (F2 is None) and PASSES on golden (F2 has formula)
    try:
        f2_value = ws['F2'].value
        if f2_value is None:
            print(f"FAIL: Component 1 — F2 is empty (None), expected a formula")
        elif not isinstance(f2_value, str) or not f2_value.startswith('='):
            print(f"FAIL: Component 1 — F2 is not a formula: {repr(f2_value)}")
        else:
            f2_upper = f2_value.upper().replace(' ', '')
            has_rank = 'RANK(' in f2_upper
            has_count = 'COUNT(' in f2_upper
            has_e2_ref = 'E2' in f2_upper or 'E$2' in f2_upper
            has_e201_ref = 'E201' in f2_upper or 'E$201' in f2_upper
            has_e_col_range = has_e2_ref and has_e201_ref

            if has_rank and has_count and has_e_col_range:
                print(f"PASS: Component 1 — F2 contains RANK and COUNT referencing E2:E201 range: {repr(f2_value)} (0.5 pts)")
                total_score += 0.5
            else:
                missing = []
                if not has_rank:
                    missing.append('RANK function')
                if not has_count:
                    missing.append('COUNT function')
                if not has_e_col_range:
                    missing.append('E2:E201 range reference')
                print(f"FAIL: Component 1 — F2 formula missing: {', '.join(missing)}. Found: {repr(f2_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: F2 formula uses the correct 1-RANK/COUNT structure
    # with ascending rank (0=descending) so higher scores get higher percentile
    # The formula should be: =1-RANK(E2,E$2:E$201,0)/COUNT(E$2:E$201)
    # We accept equivalent forms (with or without $ on row, different spacing)
    # (0.3 points)
    try:
        f2_value = ws['F2'].value
        if f2_value is None or not isinstance(f2_value, str):
            print(f"FAIL: Component 2 — F2 is not a formula, cannot check structure")
        else:
            f2_norm = normalize_formula(f2_value)
            # Check for 1- prefix (percentile = 1 - rank/count)
            has_one_minus = f2_norm.startswith('=1-')
            # Check rank uses order=0 (descending rank) which means higher score = lower rank number = higher percentile
            # Pattern: RANK(E2,...,0) or RANK(E2,...,0.0) etc
            has_rank_desc = bool(re.search(r'RANK\(E\$?2,E\$?2:E\$?201,0\)', f2_norm))
            # Check count uses correct range
            has_count_range = bool(re.search(r'COUNT\(E\$?2:E\$?201\)', f2_norm))

            if has_one_minus and has_rank_desc and has_count_range:
                print(f"PASS: Component 2 — F2 formula has correct 1-RANK/COUNT structure with order=0: {repr(f2_value)} (0.3 pts)")
                total_score += 0.3
            else:
                issues = []
                if not has_one_minus:
                    issues.append("formula should start with =1- (percentile = 1 - rank_position/count)")
                if not has_rank_desc:
                    issues.append("RANK should reference E2, range E$2:E$201, and order=0 (descending)")
                if not has_count_range:
                    issues.append("COUNT should reference E$2:E$201 or E2:E201")
                print(f"FAIL: Component 2 — {'; '.join(issues)}. Found: {repr(f2_value)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: F2 has a formula AND no other cells were inadvertently modified
    # (F3:F201 remain empty, E column rows 2-201 unchanged with E2=87)
    # This check FAILS on initial (F2 is None, so the compound condition fails) and
    # PASSES on golden (F2 has formula AND other cells are untouched).
    # (0.2 points)
    try:
        f2_value = ws['F2'].value
        # Gate: F2 must be a formula (fails on initial where F2=None)
        if f2_value is None or not isinstance(f2_value, str) or not f2_value.startswith('='):
            print(f"FAIL: Component 3 — F2 is not a formula ({repr(f2_value)}), cannot verify cell integrity")
        else:
            # Check E2 is still 87 (unchanged)
            e2_val = ws['E2'].value
            # Check F3:F201 are all still empty (formula only placed in F2)
            f_modified = []
            for row in range(3, 202):
                cell_val = ws.cell(row=row, column=6).value
                if cell_val is not None:
                    f_modified.append(f"F{row}={repr(cell_val)}")

            e2_ok = (e2_val == 87)
            f_clean = (len(f_modified) == 0)

            if e2_ok and f_clean:
                print(f"PASS: Component 3 — F2 has formula, E2 unchanged (={e2_val}), F3:F201 all empty (0.2 pts)")
                total_score += 0.2
            else:
                if not e2_ok:
                    print(f"FAIL: Component 3 — E2 changed, expected 87, found {repr(e2_val)}")
                if not f_clean:
                    print(f"FAIL: Component 3 — Other F-column cells modified: {f_modified[:5]}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
