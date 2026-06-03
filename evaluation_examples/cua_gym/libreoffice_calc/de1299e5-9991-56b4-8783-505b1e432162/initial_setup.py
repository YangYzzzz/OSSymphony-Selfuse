"""
Initial Setup: Performance review spreadsheet with 200 employee records
Task ID: calc_fmb_percentile_benchmarking_076
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_percentile_benchmarking_076'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Performance Review'

    # Row 1: Headers
    ws['A1'] = 'Emp ID'
    ws['B1'] = 'Name'
    ws['C1'] = 'Department'
    ws['D1'] = 'Years'
    ws['E1'] = 'Score'
    ws['F1'] = 'Percentile Rank'

    # We need 200 employees with:
    # - E2 = 87
    # - Exactly 162 employees with score < 87 (so rank of 87 from top is 39th, percentile ~0.81)
    # - Remaining 37 employees (rows 3-201 minus those 162) have score >= 87
    # - Actually: E2=87, plus 37 others with score >= 87, plus 162 with score < 87
    #   Total: 1 (E2=87) + 37 (>=87 including possibly more =87) + 162 (<87) = 200
    # Context says: 162 employees have a score BELOW 87, so 37 have score >= 87 (including E2)
    # So among rows 3-201 (199 employees): 162 have score < 87, 37 have score >= 87

    departments = [
        'Engineering', 'Marketing', 'Sales', 'Finance', 'HR',
        'Operations', 'Product', 'Legal', 'Research', 'IT Support'
    ]

    first_names = [
        'James', 'Mary', 'Robert', 'Patricia', 'John', 'Jennifer', 'Michael', 'Linda',
        'William', 'Barbara', 'David', 'Elizabeth', 'Richard', 'Susan', 'Joseph', 'Jessica',
        'Thomas', 'Sarah', 'Charles', 'Karen', 'Christopher', 'Lisa', 'Daniel', 'Nancy',
        'Matthew', 'Betty', 'Anthony', 'Margaret', 'Mark', 'Sandra', 'Donald', 'Ashley',
        'Steven', 'Dorothy', 'Paul', 'Kimberly', 'Andrew', 'Emily', 'Joshua', 'Donna',
        'Kenneth', 'Michelle', 'Kevin', 'Carol', 'Brian', 'Amanda', 'George', 'Melissa',
        'Edward', 'Deborah', 'Ronald', 'Stephanie', 'Timothy', 'Rebecca', 'Jason', 'Sharon',
        'Jeffrey', 'Laura', 'Ryan', 'Cynthia', 'Jacob', 'Kathleen', 'Gary', 'Amy',
        'Nicholas', 'Angela', 'Eric', 'Shirley', 'Jonathan', 'Anna', 'Stephen', 'Brenda',
        'Larry', 'Pamela', 'Justin', 'Emma', 'Scott', 'Nicole', 'Brandon', 'Helen',
        'Frank', 'Samantha', 'Raymond', 'Katherine', 'Gregory', 'Christine', 'Samuel', 'Debra',
        'Patrick', 'Rachel', 'Alexander', 'Carolyn', 'Jack', 'Janet', 'Dennis', 'Maria',
        'Jerry', 'Heather', 'Tyler', 'Diane', 'Aaron', 'Julie', 'Jose', 'Joyce',
        'Henry', 'Victoria'
    ]

    last_names = [
        'Smith', 'Johnson', 'Williams', 'Brown', 'Jones', 'Garcia', 'Miller', 'Davis',
        'Rodriguez', 'Martinez', 'Hernandez', 'Lopez', 'Gonzalez', 'Wilson', 'Anderson',
        'Thomas', 'Taylor', 'Moore', 'Jackson', 'Martin', 'Lee', 'Perez', 'Thompson',
        'White', 'Harris', 'Sanchez', 'Clark', 'Ramirez', 'Lewis', 'Robinson', 'Walker',
        'Young', 'Allen', 'King', 'Wright', 'Scott', 'Torres', 'Nguyen', 'Hill', 'Flores',
        'Green', 'Adams', 'Nelson', 'Baker', 'Hall', 'Rivera', 'Campbell', 'Mitchell',
        'Carter', 'Roberts', 'Chen', 'Kim', 'Patel', 'Murphy', 'Cook', 'Rogers', 'Morgan',
        'Peterson', 'Cooper', 'Reed', 'Bailey', 'Bell', 'Gomez', 'Kelly', 'Howard',
        'Ward', 'Cox', 'Diaz', 'Richardson', 'Wood', 'Watson', 'Brooks', 'Bennett',
        'Gray', 'James', 'Reyes', 'Cruz', 'Hughes', 'Price', 'Myers', 'Long', 'Foster',
        'Sanders', 'Ross', 'Morales', 'Powell', 'Sullivan', 'Russell', 'Ortiz', 'Jenkins',
        'Gutierrez', 'Perry', 'Butler', 'Barnes', 'Fisher', 'Henderson', 'Coleman', 'Simmons',
        'Patterson', 'Jordan', 'Reynolds', 'Hamilton'
    ]

    # Build scores array: E2=87, then 37 others >= 87, then 162 others < 87
    # Scores for >= 87 employees (excluding E2): 37 values in range [87, 100]
    # Scores for < 87 employees: 162 values in range [1, 86]
    import random
    random.seed(42)  # reproducibility

    scores_above = [87] * 1  # E2 itself
    # 37 more scores >= 87 for rows 3 to 39 (indices 1..37 in 0-indexed remaining)
    # Actually let's build all 200 scores: first is 87, then 37 values [87..100], then 162 values [1..86]
    # But we want exactly 162 below 87 → other 37 are >= 87
    high_scores = []
    for i in range(37):
        high_scores.append(random.randint(87, 100))

    low_scores = []
    for i in range(162):
        low_scores.append(random.randint(1, 86))

    # Interleave: E2=87 fixed, then shuffle the rest
    rest_scores = high_scores + low_scores
    random.shuffle(rest_scores)

    # Employee data rows 2-201
    for i in range(200):
        row = i + 2  # rows 2 to 201
        emp_id = f'EMP{1000 + i:04d}'

        fn_idx = i % len(first_names)
        ln_idx = (i * 3 + 7) % len(last_names)
        name = f'{first_names[fn_idx]} {last_names[ln_idx]}'

        dept = departments[i % len(departments)]
        years = (i % 20) + 1  # 1 to 20 years

        if row == 2:
            score = 87  # E2 must be 87
        else:
            score = rest_scores[i - 1]  # i-1 because row 2 (i=0) is fixed

        ws.cell(row=row, column=1, value=emp_id)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=dept)
        ws.cell(row=row, column=4, value=years)
        ws.cell(row=row, column=5, value=score)
        # Column F (Percentile Rank) is intentionally left empty

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet: Performance Review')
    print(f'Rows: 201 (1 header + 200 data rows)')
    print(f'E2 = 87 (fixed)')
    print(f'F2 = empty (target cell)')


create_initial()
