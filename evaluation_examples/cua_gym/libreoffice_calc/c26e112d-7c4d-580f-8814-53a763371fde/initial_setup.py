"""
Initial Setup: Project Change Request Log
Task ID: calc_ops_project_change_log_053
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_project_change_log_053'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: ChangeLog ---
    ws = wb.active
    ws.title = 'ChangeLog'

    # Headers in row 1
    headers = [
        'CR Number', 'Date Raised', 'Description',
        'Change Type', 'Requestor', 'Budget Impact $',
        'Impact Level', 'Status', 'Approved By'
    ]
    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    # 30 Change Requests — realistic content
    # D (Change Type), G (Impact Level), H (Status) intentionally left blank (dropdowns to be added)
    change_requests = [
        # CR Number, Date Raised, Description, (blank D), Requestor, Budget Impact $, (blank G), (blank H), Approved By
        ('CR-001', '2025-01-08', 'Expand data migration scope to include legacy records', '', 'Jennifer Walsh', 45000, '', '', ''),
        ('CR-002', '2025-01-14', 'Extend project timeline by 3 weeks due to vendor delays', '', 'Marcus Thompson', -5000, '', '', ''),
        ('CR-003', '2025-01-20', 'Add two additional backend engineers for sprint 4', '', 'Sarah Patel', 62000, '', '', ''),
        ('CR-004', '2025-01-27', 'Replace on-premise database server with cloud solution', '', 'David Kim', 18500, '', '', ''),
        ('CR-005', '2025-02-03', 'Reduce contractor hours for UI development phase', '', 'Rachel Nguyen', -22000, '', '', ''),
        ('CR-006', '2025-02-10', 'Include mobile app module in project deliverables', '', 'Tyler Brooks', 95000, '', '', ''),
        ('CR-007', '2025-02-14', 'Shift integration testing phase to overlap with development', '', 'Emily Carter', 3500, '', '', ''),
        ('CR-008', '2025-02-20', 'Engage specialist consultant for security audit', '', 'James O\'Brien', 28000, '', '', ''),
        ('CR-009', '2025-02-26', 'Update API specifications based on new regulatory requirements', '', 'Linda Fernandez', 11000, '', '', ''),
        ('CR-010', '2025-03-04', 'Add automated testing framework to CI/CD pipeline', '', 'Nathan Edwards', 15500, '', '', ''),
        ('CR-011', '2025-03-10', 'Consolidate three reporting modules into single dashboard', '', 'Olivia Martinez', -8000, '', '', ''),
        ('CR-012', '2025-03-17', 'Extend UAT phase by two weeks for additional testing', '', 'Brandon Lee', 12000, '', '', ''),
        ('CR-013', '2025-03-24', 'Add disaster recovery configuration for primary database', '', 'Chloe Robinson', 34000, '', '', ''),
        ('CR-014', '2025-03-28', 'Reduce print module scope due to business priority shift', '', 'Aaron Scott', -16000, '', '', ''),
        ('CR-015', '2025-04-02', 'Integrate with third-party payment processor', '', 'Megan Harris', 42000, '', '', ''),
        ('CR-016', '2025-04-09', 'Accelerate deployment schedule by 2 weeks', '', 'Derek Wilson', 7500, '', '', ''),
        ('CR-017', '2025-04-14', 'Hire additional QA tester for regression testing', '', 'Natalie Brown', 19000, '', '', ''),
        ('CR-018', '2025-04-21', 'Implement SSO authentication across all modules', '', 'Patrick Young', 25500, '', '', ''),
        ('CR-019', '2025-04-28', 'Upgrade server infrastructure to meet performance SLAs', '', 'Samantha Allen', 48000, '', '', ''),
        ('CR-020', '2025-05-05', 'Reduce cloud storage tier to lower monthly costs', '', 'Kevin Turner', -13500, '', '', ''),
        ('CR-021', '2025-05-12', 'Add real-time analytics feature to executive dashboard', '', 'Diana Cooper', 67000, '', '', ''),
        ('CR-022', '2025-05-19', 'Extend vendor support contract by one year', '', 'Brian Phillips', 22000, '', '', ''),
        ('CR-023', '2025-05-23', 'Reschedule go-live date due to regulatory approval delays', '', 'Christina Evans', 4000, '', '', ''),
        ('CR-024', '2025-05-30', 'Add multi-language support for international users', '', 'Robert Collins', 58000, '', '', ''),
        ('CR-025', '2025-06-04', 'Revise data retention policy and update storage strategy', '', 'Laura Stewart', 9500, '', '', ''),
        ('CR-026', '2025-06-11', 'Remove legacy reporting tool from project scope', '', 'Jason Morris', -31000, '', '', ''),
        ('CR-027', '2025-06-18', 'Onboard two additional project managers for final phase', '', 'Michelle Rogers', 53000, '', '', ''),
        ('CR-028', '2025-06-25', 'Deploy load balancing solution for production environment', '', 'Christopher Reed', 27500, '', '', ''),
        ('CR-029', '2025-07-02', 'Replace custom PDF generator with licensed third-party tool', '', 'Amanda Cook', 14000, '', '', ''),
        ('CR-030', '2025-07-09', 'Add compliance logging module for audit trail', '', 'Stephen Bailey', 37500, '', '', ''),
    ]

    for r, row_data in enumerate(change_requests, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set budget impact column (F) as currency-style number format
    for r in range(2, 32):
        ws.cell(row=r, column=6).number_format = '#,##0.00'

    # Column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 18

    # Budget summary area in K column (no formulas yet — task will add them)
    ws['K1'] = 'Original Budget'
    ws['K1'].font = Font(bold=True)
    ws['K2'] = 5000000
    ws['K2'].number_format = '$#,##0.00'
    ws['K3'] = 'Approved Changes Total'
    ws['K3'].font = Font(bold=True)
    # K4 left empty — task requires formula here
    ws['K5'] = 'Revised Budget'
    ws['K5'].font = Font(bold=True)
    # K6 left empty — task requires formula here

    ws.column_dimensions['K'].width = 24

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
