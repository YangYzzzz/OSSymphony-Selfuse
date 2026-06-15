"""
Reward Script: Project Change Request Log Setup
Task ID: calc_ops_project_change_log_053
Domain: libreoffice_calc
Scoring:
  - Component 1: Three data validation dropdowns (D, G, H columns) — 0.4 pts
  - Component 2: SUMIF formula in K4 for approved budget impacts — 0.2 pts
  - Component 3: Revised budget formula in K6 (=K2+K4) — 0.2 pts
  - Component 4: Conditional formatting on H column (Approved=green, Rejected=red, Under Review=orange) — 0.2 pts
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_ops_project_change_log_053'


def normalize_formula(formula):
    """Normalize a formula string for comparison: uppercase, remove whitespace."""
    if not isinstance(formula, str):
        return ''
    return formula.upper().replace(' ', '').replace('"', '"').replace('"', '"')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: ChangeLog sheet must exist
    if 'ChangeLog' not in wb.sheetnames:
        print("CRITICAL: 'ChangeLog' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['ChangeLog']

    # Component 1: Three data validation dropdowns (0.4 points total — 0.133 each)
    # - D2:D31: Change Type dropdown
    # - G2:G31: Impact Level dropdown
    # - H2:H31: Status dropdown
    try:
        validations = ws.data_validations.dataValidation
        dv_dict = {}
        for dv in validations:
            if dv.type == 'list':
                # Map sqref string to formula
                sqref_str = str(dv.sqref)
                dv_dict[sqref_str] = dv.formula1

        # Check Change Type dropdown on D2:D31
        change_type_found = False
        change_type_values = ['Scope Change', 'Schedule Change', 'Resource Change', 'Technical Change', 'Cost Change']
        for sqref, formula in dv_dict.items():
            if 'D2' in sqref and 'D31' in sqref:
                formula_clean = formula.strip('"').strip("'")
                for val in change_type_values:
                    if val in formula_clean:
                        change_type_found = True
                        break

        if change_type_found:
            print(f"PASS: Component 1a — Change Type dropdown on D2:D31 found (0.133 pts)")
            total_score += 0.133
        else:
            print(f"FAIL: Component 1a — Change Type dropdown on D2:D31 not found or missing values. Found: {dv_dict}")

        # Check Impact Level dropdown on G2:G31
        impact_found = False
        impact_values = ['High', 'Medium', 'Low']
        for sqref, formula in dv_dict.items():
            if 'G2' in sqref and 'G31' in sqref:
                formula_clean = formula.strip('"').strip("'")
                for val in impact_values:
                    if val in formula_clean:
                        impact_found = True
                        break

        if impact_found:
            print(f"PASS: Component 1b — Impact Level dropdown on G2:G31 found (0.133 pts)")
            total_score += 0.133
        else:
            print(f"FAIL: Component 1b — Impact Level dropdown on G2:G31 not found. Found: {dv_dict}")

        # Check Status dropdown on H2:H31
        status_found = False
        status_values = ['Under Review', 'Approved', 'Rejected', 'Deferred']
        for sqref, formula in dv_dict.items():
            if 'H2' in sqref and 'H31' in sqref:
                formula_clean = formula.strip('"').strip("'")
                for val in status_values:
                    if val in formula_clean:
                        status_found = True
                        break

        if status_found:
            print(f"PASS: Component 1c — Status dropdown on H2:H31 found (0.134 pts)")
            total_score += 0.134
        else:
            print(f"FAIL: Component 1c — Status dropdown on H2:H31 not found. Found: {dv_dict}")

    except Exception as e:
        print(f"ERROR: Component 1 (dropdowns) — {e}")

    # Component 2: SUMIF formula in K4 for approved budget impacts (0.2 points)
    # Expected: =SUMIF(H2:H31,"Approved",F2:F31)
    try:
        k4_value = ws['K4'].value
        if k4_value is not None and isinstance(k4_value, str):
            k4_norm = normalize_formula(k4_value)
            # Check that it's a SUMIF referencing H2:H31, "Approved", and F2:F31
            if ('SUMIF' in k4_norm and
                    'H2:H31' in k4_norm.replace(' ', '') and
                    'F2:F31' in k4_norm.replace(' ', '') and
                    'APPROVED' in k4_norm):
                print(f"PASS: Component 2 — SUMIF formula in K4: {k4_value} (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 2 — K4 formula '{k4_value}' does not match expected SUMIF(H2:H31,\"Approved\",F2:F31) pattern")
        else:
            print(f"FAIL: Component 2 — K4 is empty or not a formula, found: {repr(k4_value)}")
    except Exception as e:
        print(f"ERROR: Component 2 (K4 SUMIF formula) — {e}")

    # Component 3: Revised budget formula in K6 (0.2 points)
    # Expected: =K2+K4
    try:
        k6_value = ws['K6'].value
        if k6_value is not None and isinstance(k6_value, str):
            k6_norm = normalize_formula(k6_value)
            # Accept =K2+K4 or equivalent expressions
            if re.search(r'=?K2\+K4|=?K4\+K2', k6_norm):
                print(f"PASS: Component 3 — Revised budget formula in K6: {k6_value} (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 3 — K6 formula '{k6_value}' does not match expected =K2+K4 pattern")
        else:
            print(f"FAIL: Component 3 — K6 is empty or not a formula, found: {repr(k6_value)}")
    except Exception as e:
        print(f"ERROR: Component 3 (K6 revised budget formula) — {e}")

    # Component 4: Conditional formatting on H column (0.2 points)
    # Expected: green for Approved, red for Rejected, orange for Under Review (on H2:H31)
    try:
        cf_rules_found = list(ws.conditional_formatting)
        h_col_cf_found = False
        approved_green = False
        rejected_red = False
        under_review_orange = False

        for cf in ws.conditional_formatting:
            cf_str = str(cf)
            if 'H' in cf_str:
                for rule in ws.conditional_formatting[cf]:
                    rule_formulas = getattr(rule, 'formula', []) or []
                    rule_formulas_str = ' '.join(str(f) for f in rule_formulas).upper()
                    dxf = getattr(rule, 'dxf', None)
                    fill_color = None
                    if dxf and dxf.fill and dxf.fill.fgColor:
                        try:
                            fill_color = dxf.fill.fgColor.rgb
                        except Exception:
                            fill_color = None

                    # Check Approved -> green
                    if 'APPROVED' in rule_formulas_str and fill_color:
                        # Various greens: FF00B050, FF00FF00, etc.
                        if fill_color and fill_color.upper().startswith('FF') and (
                            '00B050' in fill_color.upper() or
                            '00FF00' in fill_color.upper() or
                            '70AD47' in fill_color.upper() or
                            'green' in str(fill_color).lower()
                        ):
                            approved_green = True
                            h_col_cf_found = True

                    # Check Rejected -> red
                    if 'REJECTED' in rule_formulas_str and fill_color:
                        # Various reds: FFFF0000, FFFF0000, etc.
                        if fill_color and (
                            'FF0000' in fill_color.upper() or
                            'FF0000' in fill_color.upper()
                        ):
                            rejected_red = True
                            h_col_cf_found = True

                    # Check Under Review -> orange
                    if 'UNDER REVIEW' in rule_formulas_str or 'UNDERREVIEW' in rule_formulas_str or 'UNDER' in rule_formulas_str:
                        if fill_color:
                            # Oranges: FFFFC000, FFFF9900, etc.
                            if 'FFC000' in fill_color.upper() or 'FF9900' in fill_color.upper() or 'ED7D31' in fill_color.upper():
                                under_review_orange = True
                                h_col_cf_found = True

        cf_conditions_met = sum([approved_green, rejected_red, under_review_orange])
        if cf_conditions_met >= 2:
            print(f"PASS: Component 4 — Conditional formatting on H column found "
                  f"(approved_green={approved_green}, rejected_red={rejected_red}, under_review_orange={under_review_orange}) (0.2 pts)")
            total_score += 0.2
        elif cf_conditions_met == 1:
            print(f"PARTIAL: Component 4 — Only 1 of 3 CF conditions met "
                  f"(approved_green={approved_green}, rejected_red={rejected_red}, under_review_orange={under_review_orange})")
            # Partial credit: at least some CF on H column was added
            total_score += 0.05
        else:
            # Check if any CF exists on H column (partial credit for attempting)
            if h_col_cf_found:
                print(f"PARTIAL: Component 4 — CF on H column present but colors don't match. "
                      f"CF ranges: {[str(cf) for cf in cf_rules_found]}")
                total_score += 0.05
            else:
                print(f"FAIL: Component 4 — No conditional formatting found on H column. "
                      f"CF ranges: {[str(cf) for cf in cf_rules_found]}")

    except Exception as e:
        print(f"ERROR: Component 4 (conditional formatting) — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
