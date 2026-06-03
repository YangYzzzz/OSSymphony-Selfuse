"""
Reward Script: Delete rows 8, 9, and 10 which contain duplicate entries
Task ID: calc_gfl_018
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.3): Total row count is 32 (header + 31 data rows)
  - Component 2 (0.3): Row 8 now contains shifted data (Anna Kowalski, C-1007)
  - Component 3 (0.2): No duplicate Customer IDs exist in the spreadsheet
  - Component 4 (0.2): Last data row (32) has correct data (Laura Fernandez, C-1031)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_018'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Customers' sheet must exist
    if 'Customers' not in wb.sheetnames:
        print("CRITICAL: 'Customers' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Customers']

    # Component 1: Total row count is 32 (header + 31 data rows) (0.3 points)
    # Initial has 35 rows; after deleting 3 duplicate rows, should have 32.
    try:
        max_row = ws.max_row
        if max_row == 32:
            print(f"PASS: Component 1 - Row count is 32 as expected (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 - Expected 32 rows, found {max_row}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Row 8 now contains shifted data (Anna Kowalski, C-1007) (0.3 points)
    # In initial_env, row 8 was a duplicate (Elena Rodriguez, C-1004).
    # After deletion, row 11 (Anna Kowalski, C-1007) should shift up to row 8.
    try:
        row8_id = ws.cell(row=8, column=1).value
        row8_name = ws.cell(row=8, column=2).value
        if row8_id == 'C-1007' and row8_name == 'Anna Kowalski':
            print(f"PASS: Component 2 - Row 8 has C-1007/Anna Kowalski after shift (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 - Row 8 has ID='{row8_id}', Name='{row8_name}'; "
                  f"expected ID='C-1007', Name='Anna Kowalski'")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: No duplicate Customer IDs (0.2 points)
    # The initial file had C-1004, C-1005, C-1006 appearing twice each.
    # After deletion, each Customer ID should appear exactly once.
    try:
        customer_ids = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            val = row[0].value
            if val is not None:
                customer_ids.append(str(val))

        from collections import Counter
        id_counts = Counter(customer_ids)
        duplicates = {cid: count for cid, count in id_counts.items() if count > 1}

        if len(duplicates) == 0:
            print(f"PASS: Component 3 - No duplicate Customer IDs found (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 - Found duplicate IDs: {duplicates}")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Last data row (32) has correct data (Laura Fernandez, C-1031) (0.2 points)
    # This verifies the bottom of the data shifted correctly.
    try:
        last_id = ws.cell(row=32, column=1).value
        last_name = ws.cell(row=32, column=2).value
        if last_id == 'C-1031' and last_name == 'Laura Fernandez':
            print(f"PASS: Component 4 - Row 32 has C-1031/Laura Fernandez (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 4 - Row 32 has ID='{last_id}', Name='{last_name}'; "
                  f"expected ID='C-1031', Name='Laura Fernandez'")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
