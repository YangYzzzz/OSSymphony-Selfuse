"""
Initial Setup: Delete duplicate rows 8, 9, 10 from customer spreadsheet
Task ID: calc_gfl_018
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_018'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Customers'

    # Headers
    headers = ['Customer ID', 'Name', 'Email', 'Phone', 'City', 'Country', 'Signup Date']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 34 customer records in rows 2-35
    # Rows 8, 9, 10 are duplicates of rows 5, 6, 7 (accidentally re-imported)
    customers = [
        # Rows 2-7: first 6 unique customers
        ['C-1001', 'Sarah Chen', 'sarah.chen@brightmail.com', '+1-415-555-0123', 'San Francisco', 'USA', '2024-01-15'],
        ['C-1002', 'Marcus Johnson', 'marcus.j@outlook.net', '+1-212-555-0456', 'New York', 'USA', '2024-02-03'],
        ['C-1003', 'Priya Sharma', 'priya.sharma@techcorp.in', '+91-98765-43210', 'Mumbai', 'India', '2024-02-18'],
        ['C-1004', 'Elena Rodriguez', 'elena.r@correo.es', '+34-611-555-789', 'Barcelona', 'Spain', '2024-03-05'],
        ['C-1005', 'Yuki Tanaka', 'yuki.tanaka@jpmail.co.jp', '+81-90-5555-1234', 'Tokyo', 'Japan', '2024-03-12'],
        ['C-1006', 'Liam O\'Brien', 'liam.obrien@dublinpost.ie', '+353-87-555-6789', 'Dublin', 'Ireland', '2024-03-20'],
        # Rows 8-10: DUPLICATES of rows 5, 6, 7 (same C-1004, C-1005, C-1006)
        ['C-1004', 'Elena Rodriguez', 'elena.r@correo.es', '+34-611-555-789', 'Barcelona', 'Spain', '2024-03-05'],
        ['C-1005', 'Yuki Tanaka', 'yuki.tanaka@jpmail.co.jp', '+81-90-5555-1234', 'Tokyo', 'Japan', '2024-03-12'],
        ['C-1006', 'Liam O\'Brien', 'liam.obrien@dublinpost.ie', '+353-87-555-6789', 'Dublin', 'Ireland', '2024-03-20'],
        # Rows 11-35: remaining 25 unique customers
        ['C-1007', 'Anna Kowalski', 'anna.k@polkamail.pl', '+48-501-555-321', 'Warsaw', 'Poland', '2024-04-01'],
        ['C-1008', 'David Kim', 'david.kim@seoultech.kr', '+82-10-5555-7890', 'Seoul', 'South Korea', '2024-04-08'],
        ['C-1009', 'Fatima Al-Hassan', 'fatima.h@gulfnet.ae', '+971-50-555-4321', 'Dubai', 'UAE', '2024-04-15'],
        ['C-1010', 'Thomas Mueller', 'thomas.m@berlinpost.de', '+49-170-555-8765', 'Berlin', 'Germany', '2024-04-22'],
        ['C-1011', 'Olivia Martin', 'olivia.m@parismail.fr', '+33-6-55-55-12-34', 'Paris', 'France', '2024-05-01'],
        ['C-1012', 'Carlos Silva', 'carlos.s@brasilnet.com.br', '+55-11-5555-6789', 'Sao Paulo', 'Brazil', '2024-05-10'],
        ['C-1013', 'Mei Lin', 'mei.lin@chinatech.cn', '+86-138-5555-0001', 'Shanghai', 'China', '2024-05-18'],
        ['C-1014', 'James Wilson', 'james.w@londonmail.co.uk', '+44-7700-555-123', 'London', 'UK', '2024-05-25'],
        ['C-1015', 'Sofia Petrov', 'sofia.p@moscownet.ru', '+7-916-555-4567', 'Moscow', 'Russia', '2024-06-02'],
        ['C-1016', 'Ahmed Ibrahim', 'ahmed.i@cairopost.eg', '+20-100-555-7890', 'Cairo', 'Egypt', '2024-06-10'],
        ['C-1017', 'Isabella Rossi', 'isabella.r@milanmail.it', '+39-345-555-2345', 'Milan', 'Italy', '2024-06-18'],
        ['C-1018', 'Raj Patel', 'raj.patel@delhitech.in', '+91-99876-54321', 'New Delhi', 'India', '2024-06-25'],
        ['C-1019', 'Emma Andersson', 'emma.a@stockholmnet.se', '+46-70-555-6789', 'Stockholm', 'Sweden', '2024-07-03'],
        ['C-1020', 'Lucas Santos', 'lucas.s@lisboamail.pt', '+351-91-555-1234', 'Lisbon', 'Portugal', '2024-07-10'],
        ['C-1021', 'Chloe Dupont', 'chloe.d@lyonpost.fr', '+33-7-55-55-56-78', 'Lyon', 'France', '2024-07-18'],
        ['C-1022', 'Hiroshi Yamamoto', 'hiroshi.y@osakanet.jp', '+81-80-5555-5678', 'Osaka', 'Japan', '2024-07-25'],
        ['C-1023', 'Natasha Volkov', 'natasha.v@spbmail.ru', '+7-921-555-8901', 'St Petersburg', 'Russia', '2024-08-02'],
        ['C-1024', 'Miguel Torres', 'miguel.t@madridnet.es', '+34-622-555-345', 'Madrid', 'Spain', '2024-08-10'],
        ['C-1025', 'Grace Okafor', 'grace.o@lagospost.ng', '+234-803-555-6789', 'Lagos', 'Nigeria', '2024-08-18'],
        ['C-1026', 'Felix Weber', 'felix.w@munichmail.de', '+49-151-555-2345', 'Munich', 'Germany', '2024-08-25'],
        ['C-1027', 'Aisha Khan', 'aisha.k@karachinet.pk', '+92-300-555-7890', 'Karachi', 'Pakistan', '2024-09-01'],
        ['C-1028', 'Oscar Lindgren', 'oscar.l@gothenburgpost.se', '+46-73-555-1234', 'Gothenburg', 'Sweden', '2024-09-08'],
        ['C-1029', 'Nina Johansson', 'nina.j@uppsalamail.se', '+46-76-555-5678', 'Uppsala', 'Sweden', '2024-09-15'],
        ['C-1030', 'Chen Wei', 'chen.wei@beijingtech.cn', '+86-139-5555-2222', 'Beijing', 'China', '2024-09-22'],
        ['C-1031', 'Laura Fernandez', 'laura.f@buenosaires.ar', '+54-11-5555-3456', 'Buenos Aires', 'Argentina', '2024-09-30'],
    ]

    for r, row_data in enumerate(customers, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths
    col_widths = [12, 22, 30, 20, 16, 14, 14]
    for i, width in enumerate(col_widths):
        ws.column_dimensions[chr(65 + i)].width = width

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Total rows: {ws.max_row} (1 header + {ws.max_row - 1} data rows)')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
