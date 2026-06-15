"""
Initial Setup: Move the 'Charts' sheet from last to after 'Data'
Task ID: calc_ps_060
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_060'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # Common styles
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # --- Sheet 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = 'Summary'

    summary_headers = ['Metric', 'Q1', 'Q2', 'Q3', 'Q4', 'Annual Total']
    for col, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    summary_data = [
        ['Total Revenue', 125400, 138700, 142300, 156800, 563200],
        ['Cost of Goods Sold', 62700, 69350, 71150, 78400, 281600],
        ['Gross Profit', 62700, 69350, 71150, 78400, 281600],
        ['Operating Expenses', 31350, 34675, 35575, 39200, 140800],
        ['Net Income', 31350, 34675, 35575, 39200, 140800],
        ['Profit Margin (%)', 25.0, 25.0, 25.0, 25.0, 25.0],
        ['Employee Count', 42, 45, 47, 51, None],
        ['Customer Satisfaction', 4.2, 4.3, 4.5, 4.6, None],
        ['Market Share (%)', 12.3, 12.8, 13.1, 13.5, None],
        ['New Customers', 87, 93, 105, 112, 397],
    ]
    for r, row_data in enumerate(summary_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_summary.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c >= 2 and val is not None and isinstance(val, (int, float)):
                if isinstance(val, float) and row_data[0] in ['Profit Margin (%)', 'Customer Satisfaction', 'Market Share (%)']:
                    cell.number_format = '0.0'
                elif isinstance(val, int) and val > 1000:
                    cell.number_format = '$#,##0'

    ws_summary.column_dimensions['A'].width = 22
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws_summary.column_dimensions[col_letter].width = 14

    # --- Sheet 2: Data ---
    ws_data = wb.create_sheet('Data')

    data_headers = ['Employee ID', 'Name', 'Department', 'Position', 'Salary', 'Start Date', 'Performance Score']
    for col, h in enumerate(data_headers, 1):
        cell = ws_data.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    employee_data = [
        ['EMP001', 'Sarah Chen', 'Engineering', 'Senior Developer', 92500, '2021-03-15', 4.5],
        ['EMP002', 'Marcus Johnson', 'Marketing', 'Campaign Manager', 78000, '2020-07-22', 4.2],
        ['EMP003', 'Priya Patel', 'Engineering', 'Tech Lead', 105000, '2019-01-10', 4.8],
        ['EMP004', 'James O\'Brien', 'Sales', 'Account Executive', 68000, '2022-05-03', 3.9],
        ['EMP005', 'Aisha Williams', 'HR', 'Recruitment Specialist', 65000, '2021-11-28', 4.1],
        ['EMP006', 'David Kim', 'Finance', 'Financial Analyst', 82000, '2020-02-14', 4.3],
        ['EMP007', 'Elena Rodriguez', 'Engineering', 'Backend Developer', 88000, '2022-01-09', 4.0],
        ['EMP008', 'Michael Brown', 'Sales', 'Sales Director', 115000, '2018-06-20', 4.7],
        ['EMP009', 'Fatima Hassan', 'Marketing', 'Content Strategist', 72000, '2021-08-05', 4.4],
        ['EMP010', 'Robert Taylor', 'Finance', 'Controller', 98000, '2019-04-17', 4.6],
        ['EMP011', 'Lisa Wang', 'Engineering', 'QA Engineer', 79000, '2022-09-12', 3.8],
        ['EMP012', 'Carlos Mendez', 'HR', 'HR Director', 95000, '2018-12-01', 4.5],
        ['EMP013', 'Sophie Martin', 'Marketing', 'Brand Manager', 84000, '2020-10-23', 4.3],
        ['EMP014', 'Ahmed Ali', 'Engineering', 'DevOps Engineer', 96000, '2021-06-15', 4.2],
        ['EMP015', 'Jennifer Lee', 'Sales', 'Regional Manager', 102000, '2019-09-08', 4.6],
    ]
    for r, row_data in enumerate(employee_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_data.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 5:
                cell.number_format = '$#,##0'
            elif c == 7:
                cell.number_format = '0.0'

    ws_data.column_dimensions['A'].width = 12
    ws_data.column_dimensions['B'].width = 20
    ws_data.column_dimensions['C'].width = 14
    ws_data.column_dimensions['D'].width = 22
    ws_data.column_dimensions['E'].width = 12
    ws_data.column_dimensions['F'].width = 14
    ws_data.column_dimensions['G'].width = 18

    # --- Sheet 3: Analysis ---
    ws_analysis = wb.create_sheet('Analysis')

    analysis_headers = ['Department', 'Headcount', 'Avg Salary', 'Total Budget', 'Avg Performance', 'Budget Utilization (%)']
    for col, h in enumerate(analysis_headers, 1):
        cell = ws_analysis.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    analysis_data = [
        ['Engineering', 5, 92100, 460500, 4.26, 94.2],
        ['Marketing', 3, 78000, 234000, 4.30, 88.5],
        ['Sales', 3, 95000, 285000, 4.40, 91.3],
        ['Finance', 2, 90000, 180000, 4.45, 96.1],
        ['HR', 2, 80000, 160000, 4.30, 89.7],
    ]
    for r, row_data in enumerate(analysis_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_analysis.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 3:
                cell.number_format = '$#,##0'
            elif c == 4:
                cell.number_format = '$#,##0'
            elif c == 5:
                cell.number_format = '0.00'
            elif c == 6:
                cell.number_format = '0.0'

    # Totals row
    ws_analysis.cell(row=7, column=1, value='Total').font = Font(bold=True)
    ws_analysis.cell(row=7, column=2, value=15).font = Font(bold=True)
    ws_analysis.cell(row=7, column=4, value=1319500).font = Font(bold=True)
    ws_analysis.cell(row=7, column=4).number_format = '$#,##0'

    ws_analysis.column_dimensions['A'].width = 14
    ws_analysis.column_dimensions['B'].width = 12
    ws_analysis.column_dimensions['C'].width = 14
    ws_analysis.column_dimensions['D'].width = 14
    ws_analysis.column_dimensions['E'].width = 16
    ws_analysis.column_dimensions['F'].width = 20

    # --- Sheet 4: Charts ---
    ws_charts = wb.create_sheet('Charts')

    # Put some chart source data and a chart
    chart_headers = ['Department', 'Budget ($)', 'Headcount']
    for col, h in enumerate(chart_headers, 1):
        cell = ws_charts.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    chart_data = [
        ['Engineering', 460500, 5],
        ['Marketing', 234000, 3],
        ['Sales', 285000, 3],
        ['Finance', 180000, 2],
        ['HR', 160000, 2],
    ]
    for r, row_data in enumerate(chart_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_charts.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 2:
                cell.number_format = '$#,##0'

    ws_charts.column_dimensions['A'].width = 14
    ws_charts.column_dimensions['B'].width = 14
    ws_charts.column_dimensions['C'].width = 12

    from openpyxl.chart import BarChart, Reference
    chart = BarChart()
    chart.type = "col"
    chart.title = "Department Budget Allocation"
    chart.y_axis.title = "Budget ($)"
    chart.x_axis.title = "Department"
    data_ref = Reference(ws_charts, min_col=2, min_row=1, max_row=6)
    cats_ref = Reference(ws_charts, min_col=1, min_row=2, max_row=6)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws_charts.add_chart(chart, "E2")

    # Verify sheet order: Summary, Data, Analysis, Charts
    assert wb.sheetnames == ['Summary', 'Data', 'Analysis', 'Charts'], f"Sheet order wrong: {wb.sheetnames}"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet order: {wb.sheetnames}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
