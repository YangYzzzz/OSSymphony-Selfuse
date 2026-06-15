"""
Initial Setup: Create demographics spreadsheet with male/female age distribution data
Task ID: calc_gcp_049
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_049'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Demographics ---
    ws = wb.active
    ws.title = 'Demographics'

    # Headers
    headers = ['AgeGroup', 'Males', 'Females']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")

    # Data rows - age group distributions
    # Males as negative values, Females as positive values
    data = [
        ['0-9',   -450, 430],
        ['10-19', -520, 510],
        ['20-29', -680, 700],
        ['30-39', -720, 740],
        ['40-49', -650, 670],
        ['50-59', -580, 600],
        ['60-69', -420, 460],
        ['70-79', -280, 320],
        ['80+',   -150, 200],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 1:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")

    # Set column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
