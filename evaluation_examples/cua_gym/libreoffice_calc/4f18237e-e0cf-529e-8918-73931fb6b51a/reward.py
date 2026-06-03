"""
Reward Script: Convert pivot table results to plain data table on new sheet
Task ID: calc_pivot_093
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): PlainTable sheet exists
  Component 2 (0.25): Headers row is correct (Category, Q1, Q2, Q3, Q4, Total)
  Component 3 (0.30): Data rows have correct category names and numeric values (plain, no formulas)
  Component 4 (0.20): Grand Total row has correct values
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_093'

# Expected data for verification
EXPECTED_HEADERS = ['Category', 'Q1', 'Q2', 'Q3', 'Q4', 'Total']

EXPECTED_DATA = {
    'Electronics':     [28500, 31200, 34800, 29500, 124000],
    'Furniture':       [22000, 25600, 21400, 23000, 92000],
    'Clothing':        [18500, 32000, 27500, 20000, 98000],
    'Food & Beverage': [24000, 26500, 28000, 25500, 104000],
    'Office Supplies': [15000, 22000, 24000, 21000, 82000],
}

EXPECTED_GRAND_TOTAL = [108000, 137300, 135700, 119000, 500000]


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: PlainTable sheet exists (0.25 points)
    try:
        if 'PlainTable' in wb.sheetnames:
            print(f"PASS: Component 1 — 'PlainTable' sheet exists (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — 'PlainTable' sheet not found. Sheets: {wb.sheetnames}")
            # Without the sheet, nothing else can be checked
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print(f"REWARD: {total_score}")
        return total_score

    ws = wb['PlainTable']

    # Component 2: Headers are correct (0.25 points)
    try:
        actual_headers = []
        for col in range(1, 7):
            val = ws.cell(row=1, column=col).value
            actual_headers.append(val)

        if actual_headers == EXPECTED_HEADERS:
            print(f"PASS: Component 2 — Headers match: {actual_headers} (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — Expected headers {EXPECTED_HEADERS}, found {actual_headers}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Data rows have correct category names and values, all plain (no formulas) (0.30 points)
    try:
        data_correct = 0
        total_categories = len(EXPECTED_DATA)  # 5

        for row_idx in range(2, 7):  # rows 2-6
            cat_name = ws.cell(row=row_idx, column=1).value
            if cat_name is None:
                print(f"FAIL: Component 3 — Row {row_idx} has no category name")
                continue

            cat_str = str(cat_name).strip()
            if cat_str not in EXPECTED_DATA:
                print(f"FAIL: Component 3 — Unexpected category '{cat_str}' in row {row_idx}")
                continue

            expected_vals = EXPECTED_DATA[cat_str]
            row_ok = True
            for col_idx in range(2, 7):  # columns B-F
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value

                # Check it's not a formula
                if isinstance(val, str) and val.startswith('='):
                    print(f"FAIL: Component 3 — Cell {cell.coordinate} contains formula '{val}', expected plain value")
                    row_ok = False
                    break

                # Check numeric value
                expected = expected_vals[col_idx - 2]
                try:
                    if abs(float(val) - expected) > 1.0:
                        print(f"FAIL: Component 3 — {cell.coordinate} expected {expected}, found {val}")
                        row_ok = False
                        break
                except (TypeError, ValueError):
                    print(f"FAIL: Component 3 — {cell.coordinate} not numeric: {val}")
                    row_ok = False
                    break

            if row_ok:
                data_correct += 1

        points = 0.30 * (data_correct / total_categories)
        if data_correct == total_categories:
            print(f"PASS: Component 3 — All {total_categories} data rows correct with plain values (0.30 pts)")
        else:
            print(f"PARTIAL: Component 3 — {data_correct}/{total_categories} data rows correct ({points:.2f} pts)")
        total_score += points
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Grand Total row (row 7) has correct values (0.20 points)
    try:
        gt_label = ws.cell(row=7, column=1).value
        if gt_label is None or 'grand total' not in str(gt_label).lower():
            print(f"FAIL: Component 4 — Row 7 col A expected 'Grand Total', found '{gt_label}'")
        else:
            gt_ok = True
            for col_idx in range(2, 7):
                cell = ws.cell(row=7, column=col_idx)
                val = cell.value

                # Must not be a formula
                if isinstance(val, str) and val.startswith('='):
                    print(f"FAIL: Component 4 — {cell.coordinate} contains formula '{val}'")
                    gt_ok = False
                    break

                expected = EXPECTED_GRAND_TOTAL[col_idx - 2]
                try:
                    if abs(float(val) - expected) > 1.0:
                        print(f"FAIL: Component 4 — {cell.coordinate} expected {expected}, found {val}")
                        gt_ok = False
                        break
                except (TypeError, ValueError):
                    print(f"FAIL: Component 4 — {cell.coordinate} not numeric: {val}")
                    gt_ok = False
                    break

            if gt_ok:
                print(f"PASS: Component 4 — Grand Total row correct (0.20 pts)")
                total_score += 0.20
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
