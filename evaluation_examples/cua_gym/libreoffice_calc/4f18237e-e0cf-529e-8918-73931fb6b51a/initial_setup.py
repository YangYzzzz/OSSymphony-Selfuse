"""
Initial Setup: Convert pivot table results to plain data table
Task ID: calc_pivot_093
Domain: libreoffice_calc

Creates a workbook with:
- RawData sheet: source transaction data (5 categories x 4 quarters)
- PivotResults sheet: pivot-table-style summary with Category rows, Quarter columns, SUM of Revenue
  Grand total = 500,000
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_093'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Categories and revenue data that sum to 500,000 ---
    categories = ['Electronics', 'Furniture', 'Clothing', 'Food & Beverage', 'Office Supplies']
    # Revenue per category per quarter (Q1, Q2, Q3, Q4)
    # Designed so grand total = 500,000
    revenue = {
        'Electronics':      [28500, 31200, 34800, 29500],   # = 124,000
        'Furniture':        [22000, 25600, 21400, 23000],   # = 92,000
        'Clothing':         [18500, 32000, 27500, 20000],   # = 98,000
        'Food & Beverage':  [24000, 26500, 28000, 25500],   # = 104,000
        'Office Supplies':  [15000, 22000, 24000, 21000],   # = 82,000
    }
    # Verify: 124000 + 92000 + 98000 + 104000 + 82000 = 500,000

    # --- Sheet 1: RawData (source transactions) ---
    ws_raw = wb.active
    ws_raw.title = 'RawData'

    raw_headers = ['Transaction ID', 'Date', 'Category', 'Quarter', 'Sales Rep', 'Revenue']
    for c, h in enumerate(raw_headers, 1):
        cell = ws_raw.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    sales_reps = ['Sarah Chen', 'Marcus Johnson', 'Priya Patel', 'David Kim', 'Elena Rodriguez',
                  'James Wilson', 'Aisha Thompson', 'Robert Garcia', 'Lisa Wang', 'Michael Brown']
    quarters = ['Q1', 'Q2', 'Q3', 'Q4']
    quarter_dates = {
        'Q1': ['2025-01-15', '2025-02-20', '2025-03-10'],
        'Q2': ['2025-04-05', '2025-05-18', '2025-06-22'],
        'Q3': ['2025-07-08', '2025-08-14', '2025-09-25'],
        'Q4': ['2025-10-03', '2025-11-19', '2025-12-11'],
    }

    row_num = 2
    txn_id = 1001
    import random
    random.seed(42)

    for cat in categories:
        for qi, q in enumerate(quarters):
            total_q = revenue[cat][qi]
            # Split into 2-3 transactions per quarter
            splits = [0.4, 0.35, 0.25]
            for si, split_pct in enumerate(splits):
                amount = round(total_q * split_pct, 2)
                if si == len(splits) - 1:
                    # Adjust last split to ensure exact sum
                    amount = round(total_q - sum(round(total_q * s, 2) for s in splits[:si]), 2)
                date = quarter_dates[q][si % len(quarter_dates[q])]
                rep = random.choice(sales_reps)
                ws_raw.cell(row=row_num, column=1, value=f'TXN-{txn_id}')
                ws_raw.cell(row=row_num, column=2, value=date)
                ws_raw.cell(row=row_num, column=3, value=cat)
                ws_raw.cell(row=row_num, column=4, value=q)
                ws_raw.cell(row=row_num, column=5, value=rep)
                ws_raw.cell(row=row_num, column=6, value=amount)
                ws_raw.cell(row=row_num, column=6).number_format = '#,##0.00'
                row_num += 1
                txn_id += 1

    # Set column widths
    ws_raw.column_dimensions['A'].width = 14
    ws_raw.column_dimensions['B'].width = 12
    ws_raw.column_dimensions['C'].width = 18
    ws_raw.column_dimensions['D'].width = 10
    ws_raw.column_dimensions['E'].width = 18
    ws_raw.column_dimensions['F'].width = 14

    # --- Sheet 2: PivotResults (pivot-table-style summary) ---
    ws_pivot = wb.create_sheet('PivotResults')

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    center_align = Alignment(horizontal="center")
    currency_fmt = '#,##0.00'
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Row 1: Title
    ws_pivot.merge_cells('A1:F1')
    ws_pivot['A1'] = 'Revenue by Category and Quarter'
    ws_pivot['A1'].font = Font(size=14, bold=True)
    ws_pivot['A1'].alignment = Alignment(horizontal="center")

    # Row 2: blank spacer

    # Row 3: Headers
    pivot_headers = ['Category', 'Q1', 'Q2', 'Q3', 'Q4', 'Grand Total']
    for c, h in enumerate(pivot_headers, 1):
        cell = ws_pivot.cell(row=3, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Rows 4-8: Category data
    for ri, cat in enumerate(categories):
        row = ri + 4
        ws_pivot.cell(row=row, column=1, value=cat).border = thin_border
        ws_pivot.cell(row=row, column=1).font = Font(bold=True)
        cat_total = 0
        for qi in range(4):
            val = revenue[cat][qi]
            cell = ws_pivot.cell(row=row, column=qi + 2, value=val)
            cell.number_format = currency_fmt
            cell.alignment = center_align
            cell.border = thin_border
            cat_total += val
        cell = ws_pivot.cell(row=row, column=6, value=cat_total)
        cell.number_format = currency_fmt
        cell.alignment = center_align
        cell.border = thin_border
        cell.font = Font(bold=True)

    # Row 9: Grand Total
    grand_total_row = 9
    ws_pivot.cell(row=grand_total_row, column=1, value='Grand Total')
    ws_pivot.cell(row=grand_total_row, column=1).font = Font(bold=True)
    ws_pivot.cell(row=grand_total_row, column=1).border = thin_border

    for qi in range(4):
        col_total = sum(revenue[cat][qi] for cat in categories)
        cell = ws_pivot.cell(row=grand_total_row, column=qi + 2, value=col_total)
        cell.number_format = currency_fmt
        cell.alignment = center_align
        cell.border = thin_border
        cell.font = Font(bold=True)

    cell = ws_pivot.cell(row=grand_total_row, column=6, value=500000)
    cell.number_format = currency_fmt
    cell.alignment = center_align
    cell.border = thin_border
    cell.font = Font(bold=True)

    # Column widths
    ws_pivot.column_dimensions['A'].width = 20
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws_pivot.column_dimensions[col_letter].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
