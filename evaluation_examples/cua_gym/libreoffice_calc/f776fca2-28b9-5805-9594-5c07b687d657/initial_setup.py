"""
Initial Setup: Order Entry spreadsheet with product code column needing validation
Task ID: calc_gcv_087
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_087'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: Order Entry ---
    ws1 = wb.active
    ws1.title = 'Sheet1'

    # Headers
    ws1.cell(row=1, column=1, value='Order ID')
    ws1.cell(row=1, column=2, value='Customer')
    ws1.cell(row=1, column=3, value='Product Code')

    # 29 orders with Order IDs and Customer names, Product Code left empty
    customers = [
        'Sarah Chen', 'Marcus Johnson', 'Elena Rodriguez', 'James Park',
        'Aisha Patel', 'David Kim', 'Maria Santos', 'Robert Taylor',
        'Yuki Tanaka', 'Fatima Al-Rashid', 'Thomas Mueller', 'Priya Sharma',
        'Carlos Mendez', 'Lisa Nakamura', 'Omar Hassan', 'Jennifer Liu',
        'Andrei Volkov', 'Sophie Martin', 'Kwame Asante', 'Isabella Rossi',
        'Wei Zhang', 'Rachel Green', 'Dmitri Petrov', 'Amara Okafor',
        'Lucas Fernandez', 'Hana Yoshida', 'Benjamin Wright', 'Nadia Kowalski',
        'Gabriel Torres'
    ]

    for i in range(29):
        row = i + 2
        ws1.cell(row=row, column=1, value=f'ORD-{i + 1:03d}')
        ws1.cell(row=row, column=2, value=customers[i])
        # Column C (Product Code) intentionally left empty - no validation

    # Set column widths for readability
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 16

    # --- Sheet2: Master Product List ---
    ws2 = wb.create_sheet('Sheet2')
    ws2.cell(row=1, column=1, value='Product Code')

    # 99 valid product codes: PRD-001 through PRD-099
    for i in range(1, 100):
        ws2.cell(row=i + 1, column=1, value=f'PRD-{i:03d}')

    ws2.column_dimensions['A'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Open in LibreOffice Calc for GUI-ready state
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
