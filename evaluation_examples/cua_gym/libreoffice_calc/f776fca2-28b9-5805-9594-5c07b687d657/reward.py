"""
Reward Script: Data validation on column C referencing Sheet2 product codes
Task ID: calc_gcv_087
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Data validation exists on C2:C30
  Component 2 (0.25): Validation type is list with correct source formula
  Component 3 (0.20): Input message configured correctly
  Component 4 (0.15): Error alert configured with Stop action
  Component 5 (0.10): Dropdown is shown (showDropDown=False)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_087'


def persist_app_state(domain: str):
    """Save any unsaved LibreOffice changes before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify data validation setup on Sheet1 C2:C30.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sheet1 and Sheet2 must exist
    if 'Sheet1' not in wb.sheetnames or 'Sheet2' not in wb.sheetnames:
        print(f"CRITICAL: Required sheets missing. Found: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Sheet1']
    validations = ws.data_validations.dataValidation

    # Component 1: Data validation exists covering C2:C30 (0.30 points)
    # Initial env has 0 validations, golden has 1 covering C2:C30
    try:
        matching_dv = None
        for dv in validations:
            sqref_str = str(dv.sqref).upper().replace(' ', '')
            # Check if C2:C30 is covered by this validation range
            if 'C2:C30' in sqref_str or 'C2:C30' == sqref_str:
                matching_dv = dv
                break
            # Also accept if ranges are split but cover the same area
            # e.g., "C2:C30" could appear in various forms
        if matching_dv is not None:
            print(f"PASS: Component 1 — Data validation found on {matching_dv.sqref} (0.30 pts)")
            total_score += 0.30
        else:
            # Check if any validation covers at least part of C2:C30
            fallback_dv = None
            for dv in validations:
                sqref_str = str(dv.sqref).upper()
                if 'C' in sqref_str:
                    fallback_dv = dv
                    matching_dv = dv  # use this for subsequent checks even if range not exact
                    break
            if fallback_dv is not None:
                print(f"FAIL: Component 1 — Found validation on column C but range is {matching_dv.sqref}, expected C2:C30")
            else:
                print(f"FAIL: Component 1 — No data validation found on column C. Total validations: {len(validations)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # If no matching DV was found at all, remaining checks will all fail
    if matching_dv is None and len(validations) > 0:
        # Try the first validation as fallback for remaining checks
        for dv in validations:
            matching_dv = dv
            break

    if matching_dv is None:
        print("No data validation found — remaining components cannot be checked")
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    # Component 2: Validation type is list with source Sheet2.$A$2:$A$100 (0.25 points)
    # Initial env has no validation, so this only passes on golden
    try:
        dv_type = matching_dv.type
        formula1 = str(matching_dv.formula1) if matching_dv.formula1 else ''
        formula1_norm = formula1.upper().replace(' ', '')

        type_ok = (dv_type == 'list')
        # Accept various valid forms of the reference
        source_ok = any(
            valid_src in formula1_norm
            for valid_src in [
                'SHEET2!$A$2:$A$100',
                'SHEET2.$A$2:$A$100',
                '$SHEET2.$A$2:$A$100',
            ]
        )

        if type_ok and source_ok:
            print(f"PASS: Component 2 — type={dv_type}, formula1={formula1} (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — type={dv_type} (expected 'list'), formula1={formula1} (expected Sheet2!$A$2:$A$100)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Input message configured (title and message) (0.20 points)
    # Initial env has no validation so no input message
    try:
        show_input = matching_dv.showInputMessage
        prompt_title = matching_dv.promptTitle or ''
        prompt_msg = matching_dv.prompt or ''

        # Check that input message is enabled and has appropriate content
        has_title = len(prompt_title.strip()) > 0
        has_message = len(prompt_msg.strip()) > 0
        # Verify title contains "Product Code" (case-insensitive)
        title_matches = 'product' in prompt_title.lower() and 'code' in prompt_title.lower()
        # Verify message mentions product code selection
        msg_relevant = 'product' in prompt_msg.lower() or 'valid' in prompt_msg.lower() or 'select' in prompt_msg.lower()

        if show_input and has_title and has_message and title_matches and msg_relevant:
            print(f"PASS: Component 3 — Input help: title='{prompt_title}', msg='{prompt_msg}' (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 3 — showInputMessage={show_input}, title='{prompt_title}', msg='{prompt_msg}'")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Error alert configured with Stop action (0.15 points)
    # Initial env has no validation so no error alert
    try:
        show_error = matching_dv.showErrorMessage
        error_style = matching_dv.errorStyle or ''
        error_msg = matching_dv.error or ''
        error_title = matching_dv.errorTitle or ''

        style_ok = (error_style.lower() == 'stop')
        has_error_msg = len(error_msg.strip()) > 0

        if show_error and style_ok and has_error_msg:
            print(f"PASS: Component 4 — Error alert: style={error_style}, title='{error_title}', msg='{error_msg}' (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — showErrorMessage={show_error}, errorStyle={error_style}, error='{error_msg}'")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Dropdown is shown (showDropDown=False in openpyxl) (0.10 points)
    # Initial env has no validation, so no dropdown
    try:
        # In openpyxl, showDropDown=False means the dropdown IS shown (inverted logic)
        # We also accept None as it defaults to showing the dropdown
        show_dropdown = matching_dv.showDropDown
        if show_dropdown is False or show_dropdown is None:
            print(f"PASS: Component 5 — Dropdown shown (showDropDown={show_dropdown}) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 — showDropDown={show_dropdown} (expected False for dropdown to appear)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
