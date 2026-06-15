"""
Initial Setup: Create spreadsheet with triple metrics data for combination chart task
Task ID: calc_gcp_058
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_058'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'TripleMetrics'

    # Headers
    headers = ['Month', 'Revenue', 'Cost', 'ProfitMargin%']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Monthly data (Jan-Dec) — Revenue $80K-$150K, Cost $50K-$95K, ProfitMargin% 15%-42%
    data = [
        ['Jan',  82500,  54300, 0.342],
        ['Feb',  88900,  58700, 0.340],
        ['Mar',  95200,  61400, 0.355],
        ['Apr', 101800,  67200, 0.340],
        ['May', 108500,  72500, 0.332],
        ['Jun', 112300,  74100, 0.340],
        ['Jul', 119700,  78600, 0.344],
        ['Aug', 125400,  82300, 0.344],
        ['Sep', 131200,  85900, 0.346],
        ['Oct', 138600,  89200, 0.357],
        ['Nov', 143800,  91500, 0.364],
        ['Dec', 149500,  93800, 0.373],
    ]

    data_align = Alignment(horizontal='center', vertical='center')
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = data_align
            cell.border = thin_border
            if c == 2 or c == 3:
                cell.number_format = '$#,##0'
            elif c == 4:
                cell.number_format = '0.0%'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18

    # NO charts in initial — the task is to create the combination chart
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
