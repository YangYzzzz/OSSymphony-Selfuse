"""
Initial Setup: Create spreadsheet with student test data and a SUM pivot table.
Task ID: calc_pivot_035
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_035'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def generate_scores(n, target_avg, target_max, lo=55, hi=None):
    """Generate n integer scores with exact average and max."""
    if hi is None:
        hi = target_max
    target_sum = round(target_avg * n)
    # Start with random scores in range [lo, hi-1] (reserve room for max)
    random.seed(42 + n)  # reproducible
    scores = [random.randint(lo, hi - 5) for _ in range(n)]
    # Force the max value into the list
    max_idx = random.randint(0, n - 1)
    scores[max_idx] = target_max
    # Adjust sum to match target
    current_sum = sum(scores)
    diff = target_sum - current_sum
    # Distribute the difference across scores (avoiding the max entry)
    indices = [i for i in range(n) if i != max_idx]
    random.shuffle(indices)
    i = 0
    while diff != 0:
        idx = indices[i % len(indices)]
        step = 1 if diff > 0 else -1
        new_val = scores[idx] + step
        if lo <= new_val <= hi and new_val != target_max:
            scores[idx] = new_val
            diff -= step
        i += 1
        if i > n * 200:
            break
    return scores


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: TestData ---
    ws = wb.active
    ws.title = 'TestData'

    headers = ['ID', 'Student', 'Subject', 'Score']
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10

    # Generate realistic student data
    first_names = [
        'Sarah', 'Marcus', 'Emily', 'James', 'Olivia', 'Liam', 'Sophia', 'Noah',
        'Ava', 'William', 'Isabella', 'Ethan', 'Mia', 'Alexander', 'Charlotte',
        'Benjamin', 'Amelia', 'Daniel', 'Harper', 'Matthew', 'Evelyn', 'Lucas',
        'Abigail', 'Henry', 'Ella', 'Sebastian', 'Scarlett', 'Jack', 'Grace',
        'Owen', 'Chloe', 'Ryan', 'Lily', 'Nathan', 'Aria', 'Caleb', 'Zoey',
        'Dylan', 'Penelope', 'Andrew', 'Layla', 'Joshua', 'Riley', 'Adrian',
        'Nora', 'Gabriel', 'Luna', 'Christopher', 'Camila', 'David'
    ]
    last_names = [
        'Chen', 'Johnson', 'Williams', 'Brown', 'Garcia', 'Martinez', 'Davis',
        'Rodriguez', 'Wilson', 'Anderson', 'Taylor', 'Thomas', 'Moore', 'Jackson',
        'White', 'Harris', 'Clark', 'Lewis', 'Walker', 'Hall', 'Young', 'Allen',
        'King', 'Wright', 'Scott', 'Green', 'Baker', 'Adams', 'Nelson', 'Hill',
        'Ramirez', 'Campbell', 'Mitchell', 'Roberts', 'Carter', 'Phillips',
        'Evans', 'Turner', 'Torres', 'Parker'
    ]

    subjects_config = [
        ('Math', 70, 74.5, 98),       # (subject, count, avg, max)
        ('Science', 60, 79.2, 100),
        ('English', 70, 82.1, 97),
    ]

    random.seed(12345)
    all_rows = []
    for subject, count, avg, mx in subjects_config:
        scores = generate_scores(count, avg, mx)
        for score in scores:
            fname = random.choice(first_names)
            lname = random.choice(last_names)
            all_rows.append((f'{fname} {lname}', subject, score))

    # Shuffle to mix subjects
    random.shuffle(all_rows)

    for i, (student, subject, score) in enumerate(all_rows, 1):
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=student)
        ws.cell(row=i + 1, column=3, value=subject)
        ws.cell(row=i + 1, column=4, value=score)

    # Freeze header row
    ws.freeze_panes = 'A2'

    # --- Sheet 2: PivotSheet (manual pivot table showing SUM) ---
    ws2 = wb.create_sheet('PivotSheet')

    # Title
    ws2.merge_cells('A1:B1')
    title_cell = ws2['A1']
    title_cell.value = 'Pivot Table - Score Summary'
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center")

    # Pivot headers
    pivot_headers = ['Subject', 'SUM of Score']
    thin = Side(style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(pivot_headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_all

    # Compute SUM values from the generated data
    sums = {}
    for student, subject, score in all_rows:
        sums[subject] = sums.get(subject, 0) + score

    pivot_data = [
        ('English', sums.get('English', 0)),
        ('Math', sums.get('Math', 0)),
        ('Science', sums.get('Science', 0)),
    ]

    for r, (subj, total) in enumerate(pivot_data, 4):
        ws2.cell(row=r, column=1, value=subj).border = border_all
        c = ws2.cell(row=r, column=2, value=total)
        c.border = border_all
        c.number_format = '#,##0'

    # Grand total
    grand_total = sum(v for _, v in pivot_data)
    gt_row = 4 + len(pivot_data)
    ws2.cell(row=gt_row, column=1, value='Grand Total').font = Font(bold=True)
    ws2.cell(row=gt_row, column=1).border = border_all
    c = ws2.cell(row=gt_row, column=2, value=grand_total)
    c.font = Font(bold=True)
    c.border = border_all
    c.number_format = '#,##0'

    # Column widths
    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Verify data integrity
    counts = {}
    score_sums = {}
    maxes = {}
    for student, subject, score in all_rows:
        counts[subject] = counts.get(subject, 0) + 1
        score_sums[subject] = score_sums.get(subject, 0) + score
        maxes[subject] = max(maxes.get(subject, 0), score)
    for subj in ['Math', 'Science', 'English']:
        avg = score_sums[subj] / counts[subj]
        print(f'  {subj}: count={counts[subj]}, sum={score_sums[subj]}, avg={avg:.1f}, max={maxes[subj]}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
