"""
Reward Script: Verify pivot table changed from SUM to AVERAGE + MAX of Score
Task ID: calc_pivot_035
Domain: libreoffice_calc
Scoring:
  1. Header B3 = "AVERAGE of Score" (0.15)
  2. Header C3 = "MAX of Score" (0.15)
  3. AVERAGE values correct for 3 subjects (0.30)
  4. MAX values correct for 3 subjects (0.30)
  5. SUM column removed (no "SUM of Score" header) (0.10)
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_035'


def persist_app_state(domain: str):
    """Attempt to save any unsaved GUI edits via Ctrl+S."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check PivotSheet exists
    if 'PivotSheet' not in wb.sheetnames:
        print("FAIL: 'PivotSheet' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['PivotSheet']

    # Component 1: Header B3 contains "AVERAGE of Score" (0.15 points)
    # Initial has "SUM of Score" in B3, golden has "AVERAGE of Score"
    try:
        b3_val = ws.cell(row=3, column=2).value
        if b3_val and 'average' in str(b3_val).lower() and 'score' in str(b3_val).lower():
            print(f"PASS: Component 1 -- B3 header is '{b3_val}' (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 -- Expected 'AVERAGE of Score' in B3, found: {b3_val!r}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Header C3 contains "MAX of Score" (0.15 points)
    # Initial has no C3 header (None), golden has "MAX of Score"
    try:
        c3_val = ws.cell(row=3, column=3).value
        if c3_val and 'max' in str(c3_val).lower() and 'score' in str(c3_val).lower():
            print(f"PASS: Component 2 -- C3 header is '{c3_val}' (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 -- Expected 'MAX of Score' in C3, found: {c3_val!r}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: AVERAGE values correct for all 3 subjects (0.30 points)
    # Expected: English=82.1, Math=74.5, Science=79.2
    # These are in column B, rows 4-6. We match by subject name in column A.
    try:
        expected_avg = {'english': 82.1, 'math': 74.5, 'science': 79.2}
        avg_correct = 0
        avg_total = len(expected_avg)

        for r in range(4, ws.max_row + 1):
            subject = ws.cell(row=r, column=1).value
            if subject and str(subject).strip().lower() in expected_avg:
                actual = ws.cell(row=r, column=2).value
                exp = expected_avg[str(subject).strip().lower()]
                if actual is not None:
                    try:
                        if abs(float(actual) - exp) <= 0.15:
                            avg_correct += 1
                            print(f"  PASS: AVERAGE for {subject} = {actual} (expected ~{exp})")
                        else:
                            print(f"  FAIL: AVERAGE for {subject} = {actual} (expected ~{exp})")
                    except (ValueError, TypeError):
                        print(f"  FAIL: AVERAGE for {subject} = {actual!r} (not numeric)")
                else:
                    print(f"  FAIL: AVERAGE for {subject} is None")

        if avg_correct == avg_total:
            print(f"PASS: Component 3 -- All {avg_total} AVERAGE values correct (0.30 pts)")
            total_score += 0.30
        elif avg_correct > 0:
            partial = round(0.30 * avg_correct / avg_total, 2)
            print(f"PARTIAL: Component 3 -- {avg_correct}/{avg_total} AVERAGE values correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 -- No AVERAGE values correct")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: MAX values correct for all 3 subjects (0.30 points)
    # Expected: English=97, Math=98, Science=100
    # These are in column C, rows 4-6.
    try:
        expected_max = {'english': 97, 'math': 98, 'science': 100}
        max_correct = 0
        max_total = len(expected_max)

        for r in range(4, ws.max_row + 1):
            subject = ws.cell(row=r, column=1).value
            if subject and str(subject).strip().lower() in expected_max:
                actual = ws.cell(row=r, column=3).value
                exp = expected_max[str(subject).strip().lower()]
                if actual is not None:
                    try:
                        if abs(float(actual) - exp) <= 0.5:
                            max_correct += 1
                            print(f"  PASS: MAX for {subject} = {actual} (expected {exp})")
                        else:
                            print(f"  FAIL: MAX for {subject} = {actual} (expected {exp})")
                    except (ValueError, TypeError):
                        print(f"  FAIL: MAX for {subject} = {actual!r} (not numeric)")
                else:
                    print(f"  FAIL: MAX for {subject} is None")

        if max_correct == max_total:
            print(f"PASS: Component 4 -- All {max_total} MAX values correct (0.30 pts)")
            total_score += 0.30
        elif max_correct > 0:
            partial = round(0.30 * max_correct / max_total, 2)
            print(f"PARTIAL: Component 4 -- {max_correct}/{max_total} MAX values correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 -- No MAX values correct")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # Component 5: SUM column removed -- no header containing "SUM of Score" (0.10 points)
    # Initial has "SUM of Score" in B3. Golden should NOT have it anywhere in row 3.
    try:
        sum_headers = [
            ws.cell(row=3, column=c).value
            for c in range(1, ws.max_column + 1)
            if ws.cell(row=3, column=c).value
            and 'sum' in str(ws.cell(row=3, column=c).value).lower()
            and 'score' in str(ws.cell(row=3, column=c).value).lower()
        ]

        if len(sum_headers) == 0:
            print(f"PASS: Component 5 -- No 'SUM of Score' header found (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 -- 'SUM of Score' header still present: {sum_headers}")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
