"""
Reward Script: Summarize sales data with a pivot table
Task ID: calc_pivot_016
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20) - PivotTable sheet exists
  Component 2 (0.20) - Correct structure (Product rows, Region columns, dual data fields)
  Component 3 (0.25) - SUM of Revenue values correct (spot checks)
  Component 4 (0.20) - COUNT of OrderID values correct (spot checks)
  Component 5 (0.15) - Grand Total row correct
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_016'


def find_pivot_sheet(wb):
    """Find a sheet that looks like the pivot table (not SalesLog)."""
    for name in wb.sheetnames:
        if name.lower() != 'saleslog':
            ws = wb[name]
            # Check if it has product names in column A
            vals_a = []
            for r in range(1, min(ws.max_row + 1, 20)):
                v = ws.cell(row=r, column=1).value
                if v is not None:
                    vals_a.append(str(v).strip())
            products = {'Laptop', 'Phone', 'Tablet', 'Headphones'}
            if products.intersection(set(vals_a)):
                return ws
    return None


def find_cell_by_value(ws, target, max_row=20, max_col=20):
    """Find coordinates of a cell containing target value (case-insensitive string match)."""
    target_lower = str(target).strip().lower()
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip().lower() == target_lower:
                return (r, c)
    return None


def find_header_columns(ws, header_row=1, max_col=20):
    """Return a dict mapping header text (lowered) to column index."""
    headers = {}
    for c in range(1, max_col + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is not None:
            headers[str(v).strip().lower()] = c
    return headers


def find_product_rows(ws, product_col=1, max_row=20):
    """Return a dict mapping product name (lowered) to row index."""
    rows = {}
    for r in range(2, max_row + 1):
        v = ws.cell(row=r, column=product_col).value
        if v is not None:
            rows[str(v).strip().lower()] = r
    return rows


def val_close(actual, expected, tol=1.0):
    """Check if actual is close to expected within tolerance."""
    if actual is None:
        return False
    try:
        return abs(float(actual) - float(expected)) <= tol
    except (ValueError, TypeError):
        return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: PivotTable sheet exists (0.20 points)
    # The golden file has a second sheet named 'PivotTable' with pivot data.
    # We check for ANY sheet beyond SalesLog that contains product-based pivot data.
    try:
        pivot_ws = find_pivot_sheet(wb)
        if pivot_ws is not None:
            print(f"PASS: Component 1 — Pivot sheet found: '{pivot_ws.title}' (0.20 pts)")
            total_score += 0.20
        else:
            print("FAIL: Component 1 — No pivot table sheet found (expected a sheet with Product rows)")
            print(f"  Sheets available: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print("REWARD: 0.0")
        return 0.0

    # From here, pivot_ws is confirmed to exist. Parse its structure.
    headers = find_header_columns(pivot_ws, header_row=1, max_col=20)
    product_rows = find_product_rows(pivot_ws, product_col=1, max_row=20)

    print(f"  Headers found: {headers}")
    print(f"  Product rows found: {product_rows}")

    # Component 2: Correct structure — Product rows, Region columns, dual data fields (0.20 points)
    # Expect: 4 products (Laptop, Phone, Tablet, Headphones) as rows
    # Expect: Region columns (East, West, Central) for both SUM Revenue and COUNT OrderID
    # Expect: Both SUM and COUNT aggregations present
    try:
        required_products = {'laptop', 'phone', 'tablet', 'headphones'}
        found_products = set(product_rows.keys())
        products_ok = required_products.issubset(found_products)

        # Check for both SUM/Revenue and COUNT/OrderID-related headers
        header_text = ' '.join(headers.keys())
        has_sum_revenue = any('sum' in h and 'revenue' in h for h in headers.keys()) or \
                          any('revenue' in h for h in headers.keys())
        has_count = any('count' in h for h in headers.keys())

        # Also check for region presence in headers
        has_regions = any('east' in h for h in headers.keys()) or \
                      any('west' in h for h in headers.keys()) or \
                      any('central' in h for h in headers.keys())

        struct_score = 0.0
        if products_ok:
            struct_score += 0.08
            print(f"  Products present: {found_products & required_products}")
        else:
            print(f"  FAIL: Missing products. Found: {found_products}, need: {required_products}")

        if has_regions:
            struct_score += 0.04
            print(f"  Region columns detected")
        else:
            print(f"  FAIL: No region columns found in headers")

        if has_sum_revenue:
            struct_score += 0.04
            print(f"  SUM of Revenue data field detected")
        else:
            print(f"  FAIL: No SUM of Revenue data field found")

        if has_count:
            struct_score += 0.04
            print(f"  COUNT data field detected")
        else:
            print(f"  FAIL: No COUNT data field found")

        if struct_score > 0:
            print(f"PASS: Component 2 — Structure score: {struct_score}/0.20")
            total_score += struct_score
        else:
            print(f"FAIL: Component 2 — Pivot table structure incorrect")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: SUM of Revenue values correct (0.25 points)
    # Ground truth from context: Laptop/East revenue=28500
    # From golden VM: Laptop/Central=23083, Phone/East=38471, Grand Total Revenue=385000
    # We'll find the relevant columns dynamically
    try:
        rev_score = 0.0
        checks_passed = 0
        checks_total = 0

        # Find revenue columns by header matching
        rev_east_col = None
        rev_central_col = None
        rev_west_col = None
        rev_total_col = None

        for h, c in headers.items():
            if 'revenue' in h or 'sum' in h:
                if 'east' in h:
                    rev_east_col = c
                elif 'central' in h:
                    rev_central_col = c
                elif 'west' in h:
                    rev_west_col = c
                elif 'grand' in h or 'total' in h:
                    rev_total_col = c

        laptop_row = product_rows.get('laptop')

        # Check Laptop/East revenue = 28500
        if rev_east_col and laptop_row:
            checks_total += 1
            val = pivot_ws.cell(row=laptop_row, column=rev_east_col).value
            if val_close(val, 28500, tol=5):
                checks_passed += 1
                print(f"  Laptop/East revenue: {val} == 28500 OK")
            else:
                print(f"  Laptop/East revenue: {val} != 28500 FAIL")

        # Check Laptop/Central revenue = 23083
        if rev_central_col and laptop_row:
            checks_total += 1
            val = pivot_ws.cell(row=laptop_row, column=rev_central_col).value
            if val_close(val, 23083, tol=5):
                checks_passed += 1
                print(f"  Laptop/Central revenue: {val} == 23083 OK")
            else:
                print(f"  Laptop/Central revenue: {val} != 23083 FAIL")

        # Check Laptop/West revenue = 25647
        if rev_west_col and laptop_row:
            checks_total += 1
            val = pivot_ws.cell(row=laptop_row, column=rev_west_col).value
            if val_close(val, 25647, tol=5):
                checks_passed += 1
                print(f"  Laptop/West revenue: {val} == 25647 OK")
            else:
                print(f"  Laptop/West revenue: {val} != 25647 FAIL")

        # Check Phone/East revenue = 38471
        phone_row = product_rows.get('phone')
        if rev_east_col and phone_row:
            checks_total += 1
            val = pivot_ws.cell(row=phone_row, column=rev_east_col).value
            if val_close(val, 38471, tol=5):
                checks_passed += 1
                print(f"  Phone/East revenue: {val} == 38471 OK")
            else:
                print(f"  Phone/East revenue: {val} != 38471 FAIL")

        if checks_total > 0:
            rev_score = 0.25 * (checks_passed / checks_total)
            print(f"PASS: Component 3 — Revenue checks: {checks_passed}/{checks_total} ({rev_score:.3f} pts)")
            total_score += rev_score
        else:
            print(f"FAIL: Component 3 — Could not locate revenue columns to verify")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: COUNT of OrderID values correct (0.20 points)
    # Ground truth: Laptop/East count=22, total count=300
    try:
        count_score = 0.0
        checks_passed = 0
        checks_total = 0

        count_east_col = None
        count_central_col = None
        count_west_col = None
        count_total_col = None

        for h, c in headers.items():
            if 'count' in h:
                if 'east' in h:
                    count_east_col = c
                elif 'central' in h:
                    count_central_col = c
                elif 'west' in h:
                    count_west_col = c
                elif 'grand' in h or 'total' in h:
                    count_total_col = c

        # Check Laptop/East count = 22
        if count_east_col and laptop_row:
            checks_total += 1
            val = pivot_ws.cell(row=laptop_row, column=count_east_col).value
            if val_close(val, 22, tol=0.5):
                checks_passed += 1
                print(f"  Laptop/East count: {val} == 22 OK")
            else:
                print(f"  Laptop/East count: {val} != 22 FAIL")

        # Check Headphones/Central count = 30
        headphones_row = product_rows.get('headphones')
        if count_central_col and headphones_row:
            checks_total += 1
            val = pivot_ws.cell(row=headphones_row, column=count_central_col).value
            if val_close(val, 30, tol=0.5):
                checks_passed += 1
                print(f"  Headphones/Central count: {val} == 30 OK")
            else:
                print(f"  Headphones/Central count: {val} != 30 FAIL")

        # Check Tablet/Grand Total count = 67
        tablet_row = product_rows.get('tablet')
        if count_total_col and tablet_row:
            checks_total += 1
            val = pivot_ws.cell(row=tablet_row, column=count_total_col).value
            if val_close(val, 67, tol=0.5):
                checks_passed += 1
                print(f"  Tablet/Grand Total count: {val} == 67 OK")
            else:
                print(f"  Tablet/Grand Total count: {val} != 67 FAIL")

        if checks_total > 0:
            count_score = 0.20 * (checks_passed / checks_total)
            print(f"PASS: Component 4 — Count checks: {checks_passed}/{checks_total} ({count_score:.3f} pts)")
            total_score += count_score
        else:
            print(f"FAIL: Component 4 — Could not locate count columns to verify")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Grand Total row correct (0.15 points)
    # Grand total revenue = 385000, grand total count = 300
    try:
        gt_score = 0.0
        checks_passed = 0
        checks_total = 0

        # Find grand total row
        gt_row = None
        for key, row in product_rows.items():
            if 'grand' in key or 'total' in key:
                gt_row = row
                break
        # Also check rows beyond product rows
        if gt_row is None:
            for r in range(2, 20):
                v = pivot_ws.cell(row=r, column=1).value
                if v and 'total' in str(v).lower():
                    gt_row = r
                    break

        if gt_row is not None:
            # Check grand total revenue
            if rev_total_col:
                checks_total += 1
                val = pivot_ws.cell(row=gt_row, column=rev_total_col).value
                if val_close(val, 385000, tol=10):
                    checks_passed += 1
                    print(f"  Grand Total revenue: {val} == 385000 OK")
                else:
                    print(f"  Grand Total revenue: {val} != 385000 FAIL")

            # Check grand total count
            if count_total_col:
                checks_total += 1
                val = pivot_ws.cell(row=gt_row, column=count_total_col).value
                if val_close(val, 300, tol=0.5):
                    checks_passed += 1
                    print(f"  Grand Total count: {val} == 300 OK")
                else:
                    print(f"  Grand Total count: {val} != 300 FAIL")

            if checks_total > 0:
                gt_score = 0.15 * (checks_passed / checks_total)
                print(f"PASS: Component 5 — Grand totals: {checks_passed}/{checks_total} ({gt_score:.3f} pts)")
                total_score += gt_score
            else:
                print(f"FAIL: Component 5 — Could not find grand total columns")
        else:
            print(f"FAIL: Component 5 — No Grand Total row found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score:.3f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
