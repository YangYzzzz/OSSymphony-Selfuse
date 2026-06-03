"""
Initial Setup: Sales data for pivot table summarization
Task ID: calc_pivot_016
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_016'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def generate_revenues(count, total, lo, hi):
    """Generate `count` random integers in [lo, hi] summing exactly to `total`."""
    if count == 0:
        return []
    avg = total / count
    # Clamp lo/hi to make it feasible
    effective_lo = min(lo, int(avg * 0.7))
    effective_hi = max(hi, int(avg * 1.3))
    vals = [random.randint(effective_lo, effective_hi) for _ in range(count)]
    current = sum(vals)
    diff = total - current
    # Distribute the difference one unit at a time
    attempts = 0
    while diff != 0 and attempts < abs(diff) + count * 10:
        idx = random.randint(0, count - 1)
        if diff > 0 and vals[idx] < effective_hi:
            vals[idx] += 1
            diff -= 1
        elif diff < 0 and vals[idx] > effective_lo:
            vals[idx] -= 1
            diff += 1
        attempts += 1
    # Force remaining diff onto arbitrary entries if needed
    if diff != 0:
        for i in range(count):
            if diff == 0:
                break
            if diff > 0:
                add = min(diff, effective_hi * 2 - vals[i])
                vals[i] += add
                diff -= add
            else:
                sub = min(-diff, vals[i] - 1)
                vals[i] -= sub
                diff += sub
    return vals


def create_initial():
    random.seed(42)

    # Combo spec: (product, region) -> (count, total_revenue)
    # Constraints from task context:
    #   Laptop/East: count=22, revenue=28500
    #   Grand total: count=300, revenue=385000
    #
    # Remaining: 278 orders, 356500 revenue, avg ~1282/order
    # We distribute proportionally and adjust last combo for exact total.

    combo_spec = {
        ('Laptop', 'East'):       (22, 28500),
        ('Laptop', 'West'):       (20, 25647),
        ('Laptop', 'Central'):    (18, 23083),
        ('Phone', 'East'):        (30, 38471),
        ('Phone', 'West'):        (28, 35906),
        ('Phone', 'Central'):     (25, 32059),
        ('Tablet', 'East'):       (25, 32059),
        ('Tablet', 'West'):       (22, 28212),
        ('Tablet', 'Central'):    (20, 25647),
        ('Headphones', 'East'):   (30, 38471),
        ('Headphones', 'West'):   (30, 38471),
        ('Headphones', 'Central'): (30, 38474),
    }

    # Verify constraints
    total_count = sum(v[0] for v in combo_spec.values())
    total_rev = sum(v[1] for v in combo_spec.values())
    assert total_count == 300, f"Count mismatch: {total_count}"
    assert total_rev == 385000, f"Revenue mismatch: {total_rev}"

    quantity_ranges = {
        'Laptop': (1, 3),
        'Phone': (1, 5),
        'Tablet': (1, 4),
        'Headphones': (2, 10),
    }

    rows = []
    start_date = datetime(2025, 1, 1)

    for (product, region), (count, target_rev) in combo_spec.items():
        # Revenue per order range: centered around target_rev/count
        avg_rev = target_rev / count
        lo = max(100, int(avg_rev * 0.6))
        hi = int(avg_rev * 1.4)
        revs = generate_revenues(count, target_rev, lo, hi)
        for rev in revs:
            date = start_date + timedelta(days=random.randint(0, 364))
            qty = random.randint(*quantity_ranges[product])
            rows.append([0, date, region, product, qty, rev])

    # Shuffle and assign OrderIDs
    random.shuffle(rows)
    for i, row in enumerate(rows):
        row[0] = i + 1

    # Verify
    laptop_east_rev = sum(r[5] for r in rows if r[3] == 'Laptop' and r[2] == 'East')
    laptop_east_cnt = sum(1 for r in rows if r[3] == 'Laptop' and r[2] == 'East')
    grand_rev = sum(r[5] for r in rows)
    print(f"Total rows: {len(rows)}, Total revenue: {grand_rev}")
    print(f"Laptop/East count: {laptop_east_cnt}, revenue: {laptop_east_rev}")

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'SalesLog'

    # Headers
    headers = ['OrderID', 'Date', 'Region', 'Product', 'Quantity', 'Revenue']
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for r, row_data in enumerate(rows, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 2 and isinstance(val, datetime):
                cell.number_format = 'yyyy-mm-dd'
            elif c == 6:
                cell.number_format = '#,##0'

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
