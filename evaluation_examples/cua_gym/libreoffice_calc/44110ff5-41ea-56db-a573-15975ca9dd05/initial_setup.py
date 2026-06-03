"""
Initial Setup: Grade book with 100 students for honor roll extraction task
Task ID: calc_edu_honor_roll_extract_043
Domain: libreoffice_calc
"""

import random
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_honor_roll_extract_043'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

random.seed(42)

LAST_NAMES = [
    "Anderson", "Brown", "Campbell", "Chen", "Clark", "Davis", "Edwards",
    "Evans", "Fisher", "Garcia", "Green", "Hall", "Harris", "Jackson",
    "Johnson", "Jones", "Kim", "Lee", "Lewis", "Lopez", "Martin", "Martinez",
    "Miller", "Mitchell", "Moore", "Morgan", "Murphy", "Nelson", "Nguyen",
    "Parker", "Patel", "Perez", "Phillips", "Powell", "Reed", "Rivera",
    "Roberts", "Robinson", "Rodriguez", "Ross", "Russell", "Sanchez",
    "Scott", "Smith", "Stewart", "Taylor", "Thomas", "Thompson", "Turner",
    "Walker", "Wang", "Watson", "White", "Williams", "Wilson", "Wood",
    "Wright", "Young", "Zhang", "Adams", "Baker", "Barnes", "Bell",
    "Bennett", "Brooks", "Butler", "Carter", "Coleman", "Collins", "Cooper",
    "Cox", "Cruz", "Diaz", "Foster", "Flores", "Gonzalez", "Gray",
    "Griffin", "Hayes", "Henderson", "Hill", "Howard", "Hughes", "James",
    "Jenkins", "Kelly", "King", "Long", "Mason", "Medina", "Myers",
    "Ortiz", "Perry", "Price", "Ramirez", "Reyes", "Richardson", "Sanders",
    "Shaw", "Torres", "Tucker", "Ward", "Washington", "Wood", "Collins",
    "Howard"
]

FIRST_NAMES = [
    "Aiden", "Ava", "Benjamin", "Charlotte", "Daniel", "Emma", "Ethan",
    "Grace", "Hannah", "Isabella", "Jacob", "Jasmine", "Joshua", "Julia",
    "Kayla", "Liam", "Lily", "Lucas", "Madison", "Mason", "Mia",
    "Michael", "Nathan", "Natalie", "Noah", "Olivia", "Owen", "Sophia",
    "Tyler", "Victoria", "William", "Zoe", "Aaron", "Abigail", "Alex",
    "Alexis", "Alice", "Andrew", "Angela", "Anna", "Brandon", "Brianna",
    "Bryan", "Caleb", "Caroline", "Connor", "Crystal", "Dylan", "Elena",
    "Elijah", "Elizabeth", "Emily", "Eric", "Evan", "Fiona", "Gabriel",
    "Genesis", "Hailey", "Hunter", "Ian", "Jack", "Jennifer", "Jessica",
    "Jonathan", "Jordan", "Joseph", "Justin", "Katherine", "Kevin",
    "Kyle", "Lauren", "Logan", "Luke", "Marcus", "Maria", "Matthew",
    "Maya", "Megan", "Melanie", "Miranda", "Morgan", "Nicholas", "Nicole",
    "Patrick", "Peyton", "Rachel", "Rebecca", "Ryan", "Samantha", "Sara",
    "Sarah", "Sebastian", "Sierra", "Stephanie", "Sydney", "Timothy",
    "Vanessa", "Xavier", "Zach", "Zoey", "Andre", "Brittany"
]


def generate_letter_grade(gpa):
    """Generate a letter grade based on rough GPA, with some randomness."""
    # For high GPA (3.5+), mostly A/B
    # For mid GPA (3.0-3.5), mix of A/B/C
    # For low GPA (2.0-3.0), mix of B/C/D
    # For failing GPA (<2.0), include some F
    r = random.random()
    if gpa >= 3.5:
        if r < 0.6:
            return 'A'
        else:
            return 'B'
    elif gpa >= 3.0:
        if r < 0.4:
            return 'A'
        elif r < 0.85:
            return 'B'
        else:
            return 'C'
    elif gpa >= 2.5:
        if r < 0.15:
            return 'A'
        elif r < 0.45:
            return 'B'
        elif r < 0.80:
            return 'C'
        else:
            return 'D'
    elif gpa >= 2.0:
        if r < 0.1:
            return 'B'
        elif r < 0.35:
            return 'C'
        elif r < 0.70:
            return 'D'
        else:
            return 'F'
    else:
        if r < 0.1:
            return 'C'
        elif r < 0.3:
            return 'D'
        else:
            return 'F'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: AllStudents ---
    ws = wb.active
    ws.title = 'AllStudents'

    # Headers
    headers = ['Last Name', 'First Name', 'GPA', 'English', 'Math', 'Science', 'History']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Generate 100 students
    used_names = set()
    students = []
    i = 0
    while len(students) < 100:
        last = LAST_NAMES[i % len(LAST_NAMES)]
        first = FIRST_NAMES[(i * 3 + 7) % len(FIRST_NAMES)]
        # Avoid exact duplicates
        name_key = (last, first)
        if name_key in used_names:
            # Tweak first name with suffix
            first = first + " " + chr(65 + (i % 26))
        used_names.add((last, first))

        # Generate GPA: distribute across range
        # ~30 students with GPA >= 3.0 and no D/F (honor roll)
        # ~70 students that don't qualify
        if len(students) < 30:
            gpa = round(random.uniform(3.0, 4.0), 2)
        elif len(students) < 50:
            gpa = round(random.uniform(3.0, 3.9), 2)
        else:
            gpa = round(random.uniform(1.5, 3.4), 2)

        english = generate_letter_grade(gpa)
        math = generate_letter_grade(gpa)
        science = generate_letter_grade(gpa)
        history = generate_letter_grade(gpa)

        students.append([last, first, gpa, english, math, science, history])
        i += 1

    # Shuffle so honor roll students aren't all at the top
    random.shuffle(students)

    for row_idx, student in enumerate(students, 2):
        for col_idx, val in enumerate(student, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Print summary
    honor_count = sum(
        1 for s in students
        if s[2] >= 3.0 and 'F' not in s[3:7] and 'D' not in s[3:7]
    )
    print(f'Total students: 100')
    print(f'Honor roll qualifiers (GPA>=3.0, no D/F): {honor_count}')


create_initial()
