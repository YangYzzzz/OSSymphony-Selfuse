"""
Reward Script: Extract honor roll students to a separate sheet with formatting.
Task ID: calc_edu_honor_roll_extract_043
Domain: libreoffice_calc

Scoring Rubric (total = 1.0):
  Component 1: 'Honor Roll' sheet exists                     (0.20 pts)
  Component 2: Correct qualifying students present (68 rows) (0.25 pts)
  Component 3: Students sorted by GPA descending             (0.20 pts)
  Component 4: Header row formatting (merge, gold, bold)     (0.20 pts)
  Component 5: Borders applied on data range                 (0.15 pts)
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_honor_roll_extract_043'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # ----------------------------------------------------------------
    # Component 1: 'Honor Roll' sheet exists (0.20 points)
    # This FAILS on initial (no Honor Roll sheet) and PASSES on golden.
    # ----------------------------------------------------------------
    try:
        if 'Honor Roll' in wb.sheetnames:
            print("PASS: Component 1 — 'Honor Roll' sheet exists (0.20 pts)")
            total_score += 0.20
            ws_hr = wb['Honor Roll']
        else:
            print(f"FAIL: Component 1 — 'Honor Roll' sheet not found. Sheets: {wb.sheetnames}")
            # Without the sheet nothing else can be verified
            final_score = min(total_score, 1.0)
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {final_score}")
            return final_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    # ----------------------------------------------------------------
    # Component 2: Correct qualifying students present (0.25 points)
    # Honor roll = GPA >= 3.0, no F or D in any subject grade.
    # Expected: 68 qualifying students in rows 3-70 (row 1=title, row 2=headers).
    # Verify the student count AND that each row has a qualifying student
    # (GPA >= 3.0 and no F/D grades).
    # ----------------------------------------------------------------
    try:
        max_row = ws_hr.max_row
        # Row 1 is merged title, row 2 is column headers, rows 3+ are data
        data_rows = max_row - 2  # number of student rows
        if data_rows <= 0:
            print(f"FAIL: Component 2 — No student data rows found (max_row={max_row})")
        else:
            # Verify every student row qualifies
            all_qualify = True
            invalid_rows = []
            for r in range(3, max_row + 1):
                gpa_val = ws_hr.cell(row=r, column=3).value
                grades = [ws_hr.cell(row=r, column=c).value for c in range(4, 8)]
                try:
                    gpa_f = float(gpa_val) if gpa_val is not None else None
                except (TypeError, ValueError):
                    gpa_f = None

                if gpa_f is None or gpa_f < 3.0:
                    all_qualify = False
                    invalid_rows.append((r, 'GPA', gpa_val))
                    continue
                for g in grades:
                    if g in ('F', 'D'):
                        all_qualify = False
                        invalid_rows.append((r, 'grade', g))

            if data_rows == 68 and all_qualify:
                print(f"PASS: Component 2 — 68 qualifying students present, all meet honor roll criteria (0.25 pts)")
                total_score += 0.25
            elif data_rows == 68 and not all_qualify:
                # Count is right but some rows don't qualify
                print(f"FAIL: Component 2 — 68 rows but some are not valid honor roll: {invalid_rows[:5]}")
                # Partial: give half for correct count
                total_score += 0.10
                print(f"  Partial (0.10 pts): student count correct (68) but qualifying criteria violated")
            else:
                print(f"FAIL: Component 2 — Expected 68 students, found {data_rows}. Valid qualification: {all_qualify}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ----------------------------------------------------------------
    # Component 3: Students sorted by GPA descending (0.20 points)
    # The GPA values in column C starting at row 3 must be non-increasing.
    # ----------------------------------------------------------------
    try:
        gpas = []
        for r in range(3, ws_hr.max_row + 1):
            val = ws_hr.cell(row=r, column=3).value
            try:
                gpas.append(float(val))
            except (TypeError, ValueError):
                gpas.append(None)

        # Filter out None (shouldn't happen if component 2 passed)
        valid_gpas = [g for g in gpas if g is not None]
        is_sorted_desc = all(valid_gpas[i] >= valid_gpas[i + 1] for i in range(len(valid_gpas) - 1))

        if is_sorted_desc and len(valid_gpas) > 0:
            print(f"PASS: Component 3 — Students sorted by GPA descending (top={valid_gpas[0]}, bottom={valid_gpas[-1]}) (0.20 pts)")
            total_score += 0.20
        else:
            # Find first out-of-order pair
            out_of_order = [(i, valid_gpas[i], valid_gpas[i + 1])
                            for i in range(len(valid_gpas) - 1)
                            if valid_gpas[i] < valid_gpas[i + 1]]
            print(f"FAIL: Component 3 — GPAs not sorted descending. First violations: {out_of_order[:3]}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ----------------------------------------------------------------
    # Component 4: Header row formatting (0.20 points)
    # Row 1 must be merged A1:G1 with value 'Spring 2025 Honor Roll',
    # gold background (#FFD700 -> ARGB FFFFD700), bold, font size 14, centered.
    # ----------------------------------------------------------------
    try:
        cell_a1 = ws_hr['A1']
        passes = []
        fails = []

        # Check merged range A1:G1
        merged_ranges = [str(mr) for mr in ws_hr.merged_cells.ranges]
        if 'A1:G1' in merged_ranges:
            passes.append('merged A1:G1')
        else:
            fails.append(f'merge expected A1:G1, found {merged_ranges}')

        # Check header text
        header_val = cell_a1.value
        if header_val and 'Spring 2025 Honor Roll' in str(header_val):
            passes.append(f'title text "{header_val}"')
        else:
            fails.append(f'title expected "Spring 2025 Honor Roll", found {repr(header_val)}')

        # Check gold background (FFFFD700)
        try:
            bg_rgb = cell_a1.fill.fgColor.rgb
            if bg_rgb in ('FFFFD700', 'FFD700'):
                passes.append(f'gold fill ({bg_rgb})')
            else:
                fails.append(f'fill expected FFFFD700 (gold), found {bg_rgb}')
        except Exception as ce:
            fails.append(f'fill check error: {ce}')

        # Check bold
        if cell_a1.font.bold:
            passes.append('bold font')
        else:
            fails.append(f'font bold expected True, found {cell_a1.font.bold}')

        # Check font size 14
        font_size = cell_a1.font.size
        if font_size is not None and float(font_size) == 14.0:
            passes.append(f'font size {font_size}')
        else:
            fails.append(f'font size expected 14, found {font_size}')

        # Check horizontal centering
        align_h = cell_a1.alignment.horizontal
        if align_h == 'center':
            passes.append('centered alignment')
        else:
            fails.append(f'alignment expected center, found {align_h}')

        # Award points: if at least 4 of 6 checks pass (including title+merge as critical)
        critical_pass = ('merged A1:G1' in passes and
                         any('Spring 2025 Honor Roll' in p for p in passes))
        if len(passes) >= 5 and critical_pass:
            print(f"PASS: Component 4 — Header formatting correct: {passes} (0.20 pts)")
            total_score += 0.20
        elif len(passes) >= 3 and critical_pass:
            print(f"PARTIAL: Component 4 — Header partially formatted: PASS={passes}, FAIL={fails} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 — Header formatting missing: PASS={passes}, FAIL={fails}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ----------------------------------------------------------------
    # Component 5: Borders applied on data range (0.15 points)
    # Thin borders should be applied around/on all cells in the data range A1:G<max_row>.
    # We sample a subset: corners and middle cells.
    # ----------------------------------------------------------------
    try:
        max_row = ws_hr.max_row
        max_col = 7
        border_ok = True
        border_fails = []

        # Check a representative sample of cells (corners + first data row)
        sample_coords = [
            (1, 1), (1, max_col),           # top-left, top-right
            (max_row, 1), (max_row, max_col),  # bottom-left, bottom-right
            (2, 1), (2, max_col),             # header row
            (3, 1), (3, max_col),             # first student row
        ]

        for r, c in sample_coords:
            cell = ws_hr.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue  # skip merged cells (they inherit border from top-left)
            b = cell.border
            # Check at least one border side is not None
            has_border = any([
                b.left.style is not None,
                b.right.style is not None,
                b.top.style is not None,
                b.bottom.style is not None
            ])
            if not has_border:
                border_ok = False
                from openpyxl.utils import get_column_letter
                border_fails.append(f'{get_column_letter(c)}{r}')

        if border_ok:
            print(f"PASS: Component 5 — Borders applied on data range (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 5 — Cells without borders: {border_fails}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
