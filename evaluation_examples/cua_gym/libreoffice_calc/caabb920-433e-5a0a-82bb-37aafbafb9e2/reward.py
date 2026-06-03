"""
Reward Script: Stock portfolio tracker with gain/loss calculations, allocation pie chart, and performance formatting.
Task ID: calc_gpm_085
Domain: libreoffice_calc
Scoring:
  - Component 1: Formulas in F-I columns for rows 4-13 (0.30)
  - Component 2: TOTAL row 15 with SUM formulas and H15/G15 (0.20)
  - Component 3: Conditional formatting on H4:H13 and I4:I13 (0.20)
  - Component 4: Two charts (pie + bar) with correct titles (0.20)
  - Component 5: Number formatting on D/E/F/H ($) and I (%) columns (0.10)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_085'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Portfolio' sheet must exist
    if 'Portfolio' not in wb.sheetnames:
        print("CRITICAL: 'Portfolio' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Portfolio']

    # Component 1: Formulas in columns F-I for rows 4-13 (0.30 points)
    # Initial file has None in F-I; golden has formulas like =C4*E4, =C4*D4, =F4-G4, =H4/G4
    try:
        formula_count = 0
        expected_formulas = 40  # 10 rows x 4 columns (F, G, H, I)
        for r in range(4, 14):
            for col in [6, 7, 8, 9]:  # F, G, H, I
                val = ws.cell(row=r, column=col).value
                if val is not None and isinstance(val, str) and val.startswith('='):
                    formula_count += 1
        ratio = formula_count / expected_formulas
        if ratio >= 0.9:
            print(f"PASS: Component 1 — {formula_count}/{expected_formulas} formulas found in F-I columns (0.30 pts)")
            total_score += 0.30
        elif ratio >= 0.5:
            partial = round(0.30 * ratio, 2)
            print(f"PARTIAL: Component 1 — {formula_count}/{expected_formulas} formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — only {formula_count}/{expected_formulas} formulas found in F-I columns")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: TOTAL row 15 with SUM formulas and return ratio (0.20 points)
    # Initial file has no row 15 content; golden has TOTAL label, SUM formulas, and H15/G15
    try:
        comp2_score = 0.0

        # Check A15 has 'TOTAL' label
        a15 = ws['A15'].value
        if a15 is not None and str(a15).strip().upper() == 'TOTAL':
            comp2_score += 0.04
            print(f"PASS: Component 2a — A15 has 'TOTAL' label")
        else:
            print(f"FAIL: Component 2a — A15 expected 'TOTAL', found: {a15}")

        # Check F15 has SUM formula
        f15 = ws['F15'].value
        if f15 is not None and isinstance(f15, str) and 'SUM' in f15.upper():
            comp2_score += 0.04
            print(f"PASS: Component 2b — F15 has SUM formula: {f15}")
        else:
            print(f"FAIL: Component 2b — F15 expected SUM formula, found: {f15}")

        # Check G15 has SUM formula
        g15 = ws['G15'].value
        if g15 is not None and isinstance(g15, str) and 'SUM' in g15.upper():
            comp2_score += 0.04
            print(f"PASS: Component 2c — G15 has SUM formula: {g15}")
        else:
            print(f"FAIL: Component 2c — G15 expected SUM formula, found: {g15}")

        # Check H15 has SUM formula
        h15 = ws['H15'].value
        if h15 is not None and isinstance(h15, str) and 'SUM' in h15.upper():
            comp2_score += 0.04
            print(f"PASS: Component 2d — H15 has SUM formula: {h15}")
        else:
            print(f"FAIL: Component 2d — H15 expected SUM formula, found: {h15}")

        # Check I15 has H15/G15 formula
        i15 = ws['I15'].value
        if i15 is not None and isinstance(i15, str) and '=' in i15:
            # Should reference H15 and G15 in a division
            val_upper = i15.upper().replace(' ', '')
            if 'H15' in val_upper and 'G15' in val_upper:
                comp2_score += 0.04
                print(f"PASS: Component 2e — I15 has ratio formula: {i15}")
            else:
                print(f"FAIL: Component 2e — I15 formula doesn't reference H15/G15: {i15}")
        else:
            print(f"FAIL: Component 2e — I15 expected formula with H15/G15, found: {i15}")

        total_score += comp2_score
        print(f"Component 2 total: {comp2_score}/0.20")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Conditional formatting on H4:H13 and I4:I13 (0.20 points)
    # Initial file has no CF rules; golden has CF on both ranges
    try:
        cf_ranges = {}
        for cf in ws.conditional_formatting:
            range_str = str(cf)
            cf_ranges[range_str] = len(cf.rules)

        comp3_score = 0.0

        # Check for CF on H4:H13 (should have cellIs rules for green/red + dataBar)
        h_cf_found = False
        for range_str, rule_count in cf_ranges.items():
            if 'H4' in range_str and 'H13' in range_str:
                h_cf_found = True
                if rule_count >= 2:
                    comp3_score += 0.10
                    print(f"PASS: Component 3a — H4:H13 has {rule_count} CF rules (0.10 pts)")
                else:
                    comp3_score += 0.05
                    print(f"PARTIAL: Component 3a — H4:H13 has {rule_count} CF rule(s) (0.05 pts)")
                break

        if not h_cf_found:
            print(f"FAIL: Component 3a — No conditional formatting found for H4:H13")

        # Check for CF on I4:I13 (should have multiple cellIs rules for return % color coding)
        i_cf_found = False
        for range_str, rule_count in cf_ranges.items():
            if 'I4' in range_str and 'I13' in range_str:
                i_cf_found = True
                if rule_count >= 3:
                    comp3_score += 0.10
                    print(f"PASS: Component 3b — I4:I13 has {rule_count} CF rules (0.10 pts)")
                elif rule_count >= 1:
                    comp3_score += 0.05
                    print(f"PARTIAL: Component 3b — I4:I13 has {rule_count} CF rule(s) (0.05 pts)")
                break

        if not i_cf_found:
            print(f"FAIL: Component 3b — No conditional formatting found for I4:I13")

        total_score += comp3_score
        print(f"Component 3 total: {comp3_score}/0.20")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Two charts - pie chart + bar chart with correct titles (0.20 points)
    # Initial file has 0 charts; golden has 2
    try:
        charts = ws._charts
        comp4_score = 0.0

        if len(charts) >= 2:
            # Check for pie chart with 'Portfolio Allocation' title
            pie_found = False
            bar_found = False

            for ch in charts:
                chart_type = type(ch).__name__
                # Extract title text
                title_text = ''
                try:
                    if ch.title and hasattr(ch.title, 'tx') and ch.title.tx:
                        if hasattr(ch.title.tx, 'rich') and ch.title.tx.rich:
                            for p in ch.title.tx.rich.p:
                                for r in p.r:
                                    title_text += r.t
                except Exception:
                    pass

                if 'Pie' in chart_type:
                    pie_found = True
                    if 'Portfolio Allocation' in title_text:
                        comp4_score += 0.10
                        print(f"PASS: Component 4a — Pie chart found with title 'Portfolio Allocation' (0.10 pts)")
                    else:
                        comp4_score += 0.05
                        print(f"PARTIAL: Component 4a — Pie chart found but title='{title_text}' (0.05 pts)")

                if 'Bar' in chart_type:
                    bar_found = True
                    if 'Individual Stock Returns' in title_text:
                        comp4_score += 0.10
                        print(f"PASS: Component 4b — Bar chart found with title 'Individual Stock Returns' (0.10 pts)")
                    else:
                        comp4_score += 0.05
                        print(f"PARTIAL: Component 4b — Bar chart found but title='{title_text}' (0.05 pts)")

            if not pie_found:
                print(f"FAIL: Component 4a — No pie chart found")
            if not bar_found:
                print(f"FAIL: Component 4b — No bar chart found")
        elif len(charts) == 1:
            comp4_score += 0.05
            print(f"PARTIAL: Component 4 — Only 1 chart found, expected 2 (0.05 pts)")
        else:
            print(f"FAIL: Component 4 — No charts found (expected 2)")

        total_score += comp4_score
        print(f"Component 4 total: {comp4_score}/0.20")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Number formatting on D/E/F/H ($#,##0.00) and I (0.00%) (0.10 points)
    # Initial file likely has General format; golden has specific number formats
    try:
        comp5_score = 0.0

        # Check dollar format on D4, E4, F4, H4
        dollar_formatted = 0
        for col_letter in ['D', 'E', 'F', 'H']:
            nf = ws[f'{col_letter}4'].number_format
            if '$' in str(nf) or '#,##0' in str(nf):
                dollar_formatted += 1

        if dollar_formatted >= 3:
            comp5_score += 0.05
            print(f"PASS: Component 5a — {dollar_formatted}/4 columns have dollar formatting (0.05 pts)")
        elif dollar_formatted >= 1:
            partial = round(0.05 * dollar_formatted / 4, 2)
            comp5_score += partial
            print(f"PARTIAL: Component 5a — {dollar_formatted}/4 columns have dollar formatting ({partial} pts)")
        else:
            print(f"FAIL: Component 5a — No dollar formatting found on D/E/F/H columns")

        # Check percentage format on I4
        i_nf = ws['I4'].number_format
        if '%' in str(i_nf):
            comp5_score += 0.05
            print(f"PASS: Component 5b — I column has percentage format: {i_nf} (0.05 pts)")
        else:
            print(f"FAIL: Component 5b — I column expected % format, found: {i_nf}")

        total_score += comp5_score
        print(f"Component 5 total: {comp5_score}/0.10")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
