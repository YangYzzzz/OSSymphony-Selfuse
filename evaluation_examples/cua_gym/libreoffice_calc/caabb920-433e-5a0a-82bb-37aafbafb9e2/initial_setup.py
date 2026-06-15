"""
Initial Setup: Stock portfolio tracker with raw data only.
Task ID: calc_gpm_085
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_085'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Portfolio'

    # --- Title Row: Merge A1:I1 ---
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = 'Investment Portfolio Tracker'
    title_cell.font = Font(size=14, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='FF006400', end_color='FF006400', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Headers in Row 3 ---
    headers = ['Symbol', 'Company', 'Shares', 'Avg Cost', 'Current Price',
               'Market Value', 'Cost Basis', 'Gain/Loss', 'Return %']
    green_fill = PatternFill(start_color='FF006400', end_color='FF006400', fill_type='solid')
    white_font = Font(bold=True, color='FFFFFF')
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = white_font
        cell.fill = green_fill
        cell.alignment = center_align
        cell.border = thin_border

    # --- Stock Data Rows 4-13 ---
    # (Symbol, Company, Shares, Avg Cost, Current Price)
    # Market Value, Cost Basis, Gain/Loss, Return % left EMPTY for the task
    stocks = [
        ('AAPL',  'Apple Inc.',              150, 142.50, 178.25),
        ('MSFT',  'Microsoft Corporation',   200, 285.00, 338.50),
        ('GOOGL', 'Alphabet Inc.',            50, 120.75, 141.30),
        ('AMZN',  'Amazon.com Inc.',          80, 145.20, 152.80),
        ('NVDA',  'NVIDIA Corporation',      120, 450.00, 620.75),
        ('TSLA',  'Tesla Inc.',               60, 245.00, 215.40),
        ('META',  'Meta Platforms Inc.',       90, 290.00, 365.20),
        ('JPM',   'JPMorgan Chase & Co.',    100, 148.50, 172.90),
        ('JNJ',   'Johnson & Johnson',       110, 162.30, 155.80),
        ('VZ',    'Verizon Communications',  180, 38.75,  34.20),
    ]

    for r, (sym, comp, shares, avg_cost, cur_price) in enumerate(stocks, 4):
        ws.cell(row=r, column=1, value=sym)
        ws.cell(row=r, column=2, value=comp)
        ws.cell(row=r, column=3, value=shares)
        ws.cell(row=r, column=4, value=avg_cost)
        ws.cell(row=r, column=5, value=cur_price)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
