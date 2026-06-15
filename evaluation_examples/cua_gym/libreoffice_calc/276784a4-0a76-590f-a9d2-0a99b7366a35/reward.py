"""
Reward Script: Build a pivot table showing average test scores by subject
Task ID: calc_pivot_003
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25) - PivotTable sheet exists
  Component 2 (0.45) - Correct average values for all 4 subjects
  Component 3 (0.15) - All 4 subject names present as row labels
  Component 4 (0.15) - Grand Average row with correct value
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_003'

# Expected ground-truth averages from task context
EXPECTED_AVERAGES = {
    'math': 72.5,
    'science': 78.3,
    'english': 81.2,
    'history': 69.8,
}
EXPECTED_GRAND_AVERAGE = 75.45  # mean of the four averages


def find_pivot_sheet(wb):
    """Find a sheet that looks like a pivot table (not 'Grades')."""
    for name in wb.sheetnames:
        if name.lower() != 'grades':
            return wb[name]
    return None


def read_subject_averages(ws):
    """
    Scan the pivot sheet for subject-average pairs.
    Returns dict like {'math': 72.5, 'science': 78.3, ...}
    and the grand average value if found.
    """
    averages = {}
    grand_avg = None
    subjects = set(EXPECTED_AVERAGES.keys())

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for i, cell in enumerate(row):
            if cell.value is None:
                continue
            cell_str = str(cell.value).strip().lower()

            # Check if this cell is a subject name
            if cell_str in subjects:
                # Look for a numeric value in the next cell(s) in the same row
                for j in range(i + 1, len(row)):
                    val = row[j].value
                    if val is not None:
                        try:
                            averages[cell_str] = float(val)
                        except (ValueError, TypeError):
                            pass
                        break

            # Check for grand average row
            if 'grand' in cell_str and 'average' in cell_str:
                for j in range(i + 1, len(row)):
                    val = row[j].value
                    if val is not None:
                        try:
                            grand_avg = float(val)
                        except (ValueError, TypeError):
                            pass
                        break

    return averages, grand_avg


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: PivotTable sheet exists (0.25 points)
    # This is the primary task-introduced change: a new sheet for the pivot table
    try:
        pivot_ws = find_pivot_sheet(wb)
        if pivot_ws is not None:
            print(f"PASS: Component 1 — Pivot sheet '{pivot_ws.title}' exists (0.25 pts)")
            total_score += 0.25
        else:
            print("FAIL: Component 1 — No pivot table sheet found (only 'Grades' exists)")
            # If no pivot sheet, nothing else to check
            final_score = min(total_score, 1.0)
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {final_score}")
            return final_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    # Read pivot data once for remaining components
    averages, grand_avg = read_subject_averages(pivot_ws)
    print(f"  Found averages: {averages}")
    print(f"  Found grand average: {grand_avg}")

    # Component 2: Correct average values for all 4 subjects (0.45 points)
    # Each subject average is worth 0.45/4 = 0.1125 points
    try:
        subject_matches = 0
        for subject, expected_val in EXPECTED_AVERAGES.items():
            if subject in averages:
                actual_val = averages[subject]
                if abs(actual_val - expected_val) <= 0.15:
                    subject_matches += 1
                    print(f"  PASS: {subject.title()} average = {actual_val} (expected {expected_val})")
                else:
                    print(f"  FAIL: {subject.title()} average = {actual_val} (expected {expected_val})")
            else:
                print(f"  FAIL: {subject.title()} average not found in pivot table")

        points_per_subject = 0.45 / 4.0
        comp2_score = subject_matches * points_per_subject
        if subject_matches > 0:
            print(f"PASS: Component 2 — {subject_matches}/4 subject averages correct ({comp2_score:.4f} pts)")
            total_score += comp2_score
        else:
            print(f"FAIL: Component 2 — No subject averages matched")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: All 4 subject names present as row labels (0.15 points)
    # This checks that the pivot table has all required subjects listed
    try:
        found_subjects = set(averages.keys())
        expected_subjects = set(EXPECTED_AVERAGES.keys())
        if found_subjects >= expected_subjects:
            print(f"PASS: Component 3 — All 4 subjects present as row labels (0.15 pts)")
            total_score += 0.15
        else:
            missing = expected_subjects - found_subjects
            print(f"FAIL: Component 3 — Missing subjects: {missing}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Grand Average row with correct value (0.15 points)
    try:
        if grand_avg is not None and abs(grand_avg - EXPECTED_GRAND_AVERAGE) <= 0.2:
            print(f"PASS: Component 4 — Grand Average = {grand_avg} (expected ~{EXPECTED_GRAND_AVERAGE}) (0.15 pts)")
            total_score += 0.15
        elif grand_avg is not None:
            print(f"FAIL: Component 4 — Grand Average = {grand_avg} (expected ~{EXPECTED_GRAND_AVERAGE})")
        else:
            print(f"FAIL: Component 4 — Grand Average row not found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
