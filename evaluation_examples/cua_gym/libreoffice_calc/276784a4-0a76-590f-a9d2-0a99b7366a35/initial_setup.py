"""
Initial Setup: Build a pivot table showing average test scores by subject.
Task ID: calc_pivot_003
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_003'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def generate_scores(target_sum, count=60, low=40, high=100):
    """Generate `count` integer scores in [low, high] that sum to target_sum."""
    # Start with random scores, then adjust to hit the target sum
    random.seed(42)
    scores = [random.randint(low, high) for _ in range(count)]
    current_sum = sum(scores)
    diff = target_sum - current_sum

    # Distribute the difference across scores
    idx = 0
    while diff != 0:
        step = 1 if diff > 0 else -1
        new_val = scores[idx % count] + step
        if low <= new_val <= high:
            scores[idx % count] = new_val
            diff -= step
        idx += 1
        if idx > count * 100:
            break

    return scores


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Grades'

    # Headers
    headers = ['StudentID', 'StudentName', 'Subject', 'Score', 'Semester']
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    # Student names (60 realistic names)
    first_names = [
        'Sarah', 'Marcus', 'Emily', 'James', 'Olivia', 'Daniel', 'Sophia', 'Michael',
        'Ava', 'David', 'Isabella', 'Andrew', 'Mia', 'Christopher', 'Charlotte',
        'Matthew', 'Amelia', 'Joshua', 'Harper', 'Ryan', 'Evelyn', 'Nathan',
        'Abigail', 'Tyler', 'Ella', 'Brandon', 'Scarlett', 'Kevin', 'Grace',
        'Justin', 'Lily', 'Aaron', 'Chloe', 'Adam', 'Zoey', 'Brian',
        'Penelope', 'Jason', 'Layla', 'Patrick', 'Riley', 'Steven', 'Nora',
        'Eric', 'Hannah', 'Derek', 'Aria', 'Kyle', 'Ellie', 'Sean',
        'Stella', 'Caleb', 'Aurora', 'Trevor', 'Hazel', 'Mark', 'Luna',
        'Alex', 'Violet', 'Luke'
    ]
    last_names = [
        'Chen', 'Johnson', 'Williams', 'Garcia', 'Martinez', 'Robinson', 'Clark',
        'Lewis', 'Lee', 'Walker', 'Hall', 'Allen', 'Young', 'King', 'Wright',
        'Lopez', 'Hill', 'Scott', 'Green', 'Adams', 'Baker', 'Gonzalez',
        'Nelson', 'Carter', 'Mitchell', 'Perez', 'Roberts', 'Turner', 'Phillips',
        'Campbell', 'Parker', 'Evans', 'Edwards', 'Collins', 'Stewart', 'Sanchez',
        'Morris', 'Rogers', 'Reed', 'Cook', 'Morgan', 'Bell', 'Murphy',
        'Bailey', 'Rivera', 'Cooper', 'Richardson', 'Cox', 'Howard', 'Ward',
        'Torres', 'Peterson', 'Gray', 'Ramirez', 'James', 'Watson', 'Brooks',
        'Kelly', 'Sanders', 'Price'
    ]

    students = []
    for i in range(60):
        sid = f'S{i+1:03d}'
        name = f'{first_names[i]} {last_names[i]}'
        students.append((sid, name))

    subjects = ['Math', 'Science', 'English', 'History']
    semesters = ['Fall', 'Spring']

    # Target sums: Math=72.5*60=4350, Science=78.3*60=4698, English=81.2*60=4872, History=69.8*60=4188
    target_sums = {
        'Math': 4350,
        'Science': 4698,
        'English': 4872,
        'History': 4188
    }

    # Generate scores for each subject
    subject_scores = {}
    for subj in subjects:
        subject_scores[subj] = generate_scores(target_sums[subj], count=60, low=40, high=100)

    # Write data rows: 60 students x 4 subjects = 240 rows
    row_num = 2
    for i, (sid, name) in enumerate(students):
        for subj in subjects:
            score = subject_scores[subj][i]
            semester = semesters[i % 2]  # alternate semesters
            ws.cell(row=row_num, column=1, value=sid)
            ws.cell(row=row_num, column=2, value=name)
            ws.cell(row=row_num, column=3, value=subj)
            ws.cell(row=row_num, column=4, value=score)
            ws.cell(row=row_num, column=5, value=semester)
            row_num += 1

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Verify averages
    import openpyxl as ox
    wb2 = ox.load_workbook(OUTPUT)
    ws2 = wb2['Grades']
    sums = {'Math': 0, 'Science': 0, 'English': 0, 'History': 0}
    counts = {'Math': 0, 'Science': 0, 'English': 0, 'History': 0}
    for r in range(2, 242):
        subj = ws2.cell(row=r, column=3).value
        score = ws2.cell(row=r, column=4).value
        if subj in sums:
            sums[subj] += score
            counts[subj] += 1
    for subj in subjects:
        avg = sums[subj] / counts[subj]
        print(f'  {subj}: avg={avg:.1f} (sum={sums[subj]}, count={counts[subj]})')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
