"""
Initial Setup: Create Batch_Sizes spreadsheet with manufacturing data, no validation.
Task ID: calc_gcv_071
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_071'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Batch_Sizes"

    # Headers
    headers = ['Product', 'Line', 'Shift', 'Date', 'Batch Size']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic manufacturing data
    products = [
        'Steel Bolt M8', 'Steel Bolt M10', 'Hex Nut M8', 'Hex Nut M10',
        'Flat Washer M8', 'Spring Washer M10', 'Anchor Bolt M12',
        'Wing Nut M6', 'Lock Nut M8', 'Cap Screw M10',
        'Carriage Bolt M8', 'Eye Bolt M10', 'U-Bolt M12',
        'Stud Bolt M16', 'Coupling Nut M8', 'Flange Nut M10',
        'T-Bolt M12', 'Shoulder Bolt M8', 'Thumb Screw M6',
        'Set Screw M10',
    ]

    lines = ['Line A', 'Line B', 'Line C', 'Line D']
    shifts = ['Morning', 'Afternoon', 'Night']

    import random
    random.seed(42)

    dates_base = [
        '2025-03-03', '2025-03-04', '2025-03-05', '2025-03-06', '2025-03-07',
        '2025-03-10', '2025-03-11', '2025-03-12', '2025-03-13', '2025-03-14',
        '2025-03-17', '2025-03-18', '2025-03-19', '2025-03-20', '2025-03-21',
    ]

    for r in range(2, 51):  # rows 2-50, 49 data rows
        idx = r - 2
        product = products[idx % len(products)]
        line = lines[idx % len(lines)]
        shift = shifts[idx % len(shifts)]
        date = dates_base[idx % len(dates_base)]
        # Batch sizes that are multiples of 5 (valid production data)
        batch_size = random.choice([5, 10, 15, 20, 25, 30, 35, 40, 45, 50,
                                    55, 60, 75, 80, 100, 120, 150, 200, 250])

        ws.cell(row=r, column=1, value=product)
        ws.cell(row=r, column=2, value=line)
        ws.cell(row=r, column=3, value=shift)
        ws.cell(row=r, column=4, value=date)
        ws.cell(row=r, column=5, value=batch_size)

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
