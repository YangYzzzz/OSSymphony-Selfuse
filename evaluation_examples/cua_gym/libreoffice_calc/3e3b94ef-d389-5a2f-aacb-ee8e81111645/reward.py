"""
Reward Script: Gantt-style chart data table with formulas and bar chart
Task ID: calc_ops_041
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): D2:D7 contain End Date formulas (=BN+CN-1)
  Component 2 (0.25): Formulas reference correct cells for each row
  Component 3 (0.25): A stacked bar chart exists on the Gantt sheet
  Component 4 (0.15): Chart has exactly 2 series (Start Date base + Duration)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_041'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice document."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists
    if 'Gantt' not in wb.sheetnames:
        print("CRITICAL: 'Gantt' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Gantt']

    # =========================================================
    # Component 1: D2:D7 contain End Date formulas (0.35 points)
    # The task asks to enter formulas D = B + C - 1.
    # Initial state has D2:D7 as None, so this is a task-introduced change.
    # =========================================================
    try:
        formula_count = 0
        expected_rows = [2, 3, 4, 5, 6, 7]
        for row in expected_rows:
            cell_val = ws.cell(row=row, column=4).value  # column D
            if cell_val is not None and isinstance(cell_val, str) and cell_val.startswith('='):
                formula_count += 1
            else:
                print(f"FAIL: D{row} does not contain a formula, found: {cell_val!r}")

        if formula_count == 6:
            print(f"PASS: Component 1 -- All 6 End Date cells (D2:D7) contain formulas (0.35 pts)")
            total_score += 0.35
        elif formula_count > 0:
            partial = round(0.35 * (formula_count / 6), 2)
            print(f"PARTIAL: Component 1 -- {formula_count}/6 End Date formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 -- No formulas found in D2:D7")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # =========================================================
    # Component 2: Formulas correctly reference B+C-1 pattern (0.25 points)
    # Each formula in D{row} should reference B{row} and C{row} and subtract 1.
    # Pattern: =B{n}+C{n}-1 (allowing minor variations)
    # =========================================================
    try:
        correct_formula_count = 0
        for row in expected_rows:
            cell_val = ws.cell(row=row, column=4).value
            if cell_val is not None and isinstance(cell_val, str):
                # Normalize: uppercase, remove spaces
                normalized = cell_val.upper().replace(" ", "")
                # Accept patterns like =B2+C2-1 or =C2+B2-1
                pattern1 = f"=B{row}+C{row}-1"
                pattern2 = f"=C{row}+B{row}-1"
                if normalized == pattern1 or normalized == pattern2:
                    correct_formula_count += 1
                else:
                    print(f"FAIL: D{row} formula is '{cell_val}', expected '=B{row}+C{row}-1'")
            else:
                print(f"FAIL: D{row} has no formula to verify pattern")

        if correct_formula_count == 6:
            print(f"PASS: Component 2 -- All 6 formulas correctly reference B+C-1 (0.25 pts)")
            total_score += 0.25
        elif correct_formula_count > 0:
            partial = round(0.25 * (correct_formula_count / 6), 2)
            print(f"PARTIAL: Component 2 -- {correct_formula_count}/6 formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- No correct B+C-1 formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # =========================================================
    # Component 3: A stacked bar chart exists on the Gantt sheet (0.25 points)
    # Initial state has 0 charts, golden has 1 stacked bar chart.
    # =========================================================
    try:
        charts = ws._charts
        if len(charts) >= 1:
            # Classify charts by type and grouping
            chart_types = [ch.type for ch in charts]
            chart_groupings = [getattr(ch, 'grouping', None) for ch in charts]
            has_stacked_bar = any(
                t == 'bar' and g == 'stacked'
                for t, g in zip(chart_types, chart_groupings)
            )
            has_bar = 'bar' in chart_types

            if has_stacked_bar:
                print(f"PASS: Component 3 -- Stacked bar chart found on Gantt sheet (0.25 pts)")
                total_score += 0.25
            elif has_bar:
                print(f"PARTIAL: Component 3 -- Bar chart found but not stacked (0.15 pts)")
                total_score += 0.15
            elif len(charts) >= 1:
                print(f"PARTIAL: Component 3 -- Chart exists but is not a bar chart (type: {chart_types[0]}) (0.10 pts)")
                total_score += 0.10
        else:
            print(f"FAIL: Component 3 -- No charts found on Gantt sheet")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # =========================================================
    # Component 4: Chart has 2 series (Start Date base + Duration) (0.15 points)
    # A Gantt simulation uses Start Date as invisible base and Duration as visible bar.
    # =========================================================
    try:
        charts = ws._charts
        if len(charts) >= 1:
            # Find the bar chart
            target_chart = None
            for ch in charts:
                if ch.type == 'bar':
                    target_chart = ch
                    break
            if target_chart is None:
                target_chart = charts[0]

            series_count = len(target_chart.series)
            if series_count == 2:
                print(f"PASS: Component 4 -- Chart has exactly 2 series (0.15 pts)")
                total_score += 0.15
            elif series_count >= 1:
                print(f"PARTIAL: Component 4 -- Chart has {series_count} series, expected 2 (0.07 pts)")
                total_score += 0.07
            else:
                print(f"FAIL: Component 4 -- Chart has no data series")
        else:
            print(f"FAIL: Component 4 -- No charts to check series count")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
