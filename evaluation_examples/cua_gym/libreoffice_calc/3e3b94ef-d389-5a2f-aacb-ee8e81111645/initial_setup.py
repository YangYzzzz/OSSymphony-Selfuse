"""
Initial Setup: Build a Gantt-style chart data table for a production project
Task ID: calc_ops_041
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_041'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Gantt ---
    ws = wb.active
    ws.title = 'Gantt'

    # Headers
    headers = ['Task', 'Start Date', 'Duration (days)', 'End Date']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # Data rows — realistic production project tasks
    data = [
        ['Material Procurement', date(2026, 4, 1), 14],
        ['Quality Inspection', date(2026, 4, 15), 3],
        ['Production Run', date(2026, 4, 18), 21],
        ['Final QC', date(2026, 5, 9), 5],
        ['Packaging', date(2026, 5, 14), 7],
        ['Shipping', date(2026, 5, 21), 3],
    ]

    data_font = Font(name='Calibri', size=11)
    data_align_left = Alignment(horizontal='left', vertical='center')
    data_align_center = Alignment(horizontal='center', vertical='center')

    for r, row_data in enumerate(data, 2):
        # Task name (column A)
        cell_a = ws.cell(row=r, column=1, value=row_data[0])
        cell_a.font = data_font
        cell_a.alignment = data_align_left
        cell_a.border = header_border

        # Start Date (column B)
        cell_b = ws.cell(row=r, column=2, value=row_data[1])
        cell_b.font = data_font
        cell_b.alignment = data_align_center
        cell_b.number_format = 'yyyy-mm-dd'
        cell_b.border = header_border

        # Duration (column C)
        cell_c = ws.cell(row=r, column=3, value=row_data[2])
        cell_c.font = data_font
        cell_c.alignment = data_align_center
        cell_c.border = header_border

        # End Date (column D) — LEFT EMPTY for the agent to fill
        cell_d = ws.cell(row=r, column=4)
        cell_d.border = header_border
        cell_d.number_format = 'yyyy-mm-dd'

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 16

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
