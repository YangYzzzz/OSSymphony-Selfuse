"""
Initial Setup: VLOOKUP wildcard salary lookup
Task ID: calc_lf_010
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_010'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Payroll ---
    ws = wb.active
    ws.title = 'Payroll'

    # Headers
    headers = ['Employee ID', 'Name', 'Salary']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Employee data
    data = [
        ['EMP-039', 'Lisa Park', 72000],
        ['EMP-040', 'Mike Chen', 68000],
        ['EMP-041', 'Nina Patel', 81000],
        ['EMP-042', 'Oscar Ruiz', 75000],
        ['EMP-043', 'Priya Sen', 69000],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Salary column as currency format
    for r in range(2, 7):
        ws.cell(row=r, column=3).number_format = '$#,##0'

    # Lookup label in E1
    ws.cell(row=1, column=5, value='Salary Lookup')
    ws['E1'].font = Font(bold=True)
    ws['E1'].alignment = Alignment(horizontal='center')

    # E2 is intentionally left EMPTY (the task is to enter the formula here)

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['E'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
