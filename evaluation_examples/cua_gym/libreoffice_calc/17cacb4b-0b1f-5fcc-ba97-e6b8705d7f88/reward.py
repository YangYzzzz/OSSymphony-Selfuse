"""
Reward Script: VLOOKUP with wildcard to find salary of employee ending with "-042"
Task ID: calc_lf_010
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): E2 contains a formula
  Component 2 (0.3): Formula uses VLOOKUP with wildcard "*-042"
  Component 3 (0.3): Formula references correct range and retrieves salary (col 3)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_010'


def persist_app_state(domain: str):
    """Save any unsaved LibreOffice state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Ensure the Payroll sheet exists (precondition gate)
    if 'Payroll' not in wb.sheetnames:
        print("CRITICAL: 'Payroll' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Payroll']

    # Component 1: E2 contains a formula (0.4 points)
    # This checks the core task requirement: the agent entered a formula in E2
    try:
        e2_val = ws['E2'].value
        if e2_val is not None and isinstance(e2_val, str) and e2_val.startswith('='):
            print(f"PASS: Component 1 — E2 contains a formula: {e2_val} (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — E2 does not contain a formula, found: {repr(e2_val)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Formula uses VLOOKUP with wildcard "*-042" (0.3 points)
    # The task specifically asks for VLOOKUP with wildcard for "-042"
    try:
        e2_val = ws['E2'].value
        if e2_val is not None and isinstance(e2_val, str):
            formula_upper = e2_val.upper().replace(" ", "")
            # Check for VLOOKUP function
            has_vlookup = 'VLOOKUP(' in formula_upper
            # Check for wildcard pattern matching *-042 (with quotes)
            has_wildcard_042 = bool(re.search(r'["\'].*\*.*-042.*["\']', e2_val, re.IGNORECASE))
            if has_vlookup and has_wildcard_042:
                print(f"PASS: Component 2 — Formula uses VLOOKUP with wildcard '*-042' (0.3 pts)")
                total_score += 0.3
            else:
                if not has_vlookup:
                    print(f"FAIL: Component 2 — Formula does not use VLOOKUP: {e2_val}")
                if not has_wildcard_042:
                    print(f"FAIL: Component 2 — Formula does not contain wildcard '*-042': {e2_val}")
        else:
            print(f"FAIL: Component 2 — E2 is not a formula string: {repr(e2_val)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Formula references data range including A-C columns and retrieves column 3 (0.3 points)
    # The VLOOKUP should look up in A:C range and return column 3 (Salary)
    try:
        e2_val = ws['E2'].value
        if e2_val is not None and isinstance(e2_val, str):
            formula_upper = e2_val.upper().replace(" ", "")
            # Check that the formula references a range starting with column A and ending with column C
            # Valid patterns: A2:C6, A:C, A1:C100, etc.
            has_a_to_c_range = bool(re.search(r'A\d*:C\d*', formula_upper))
            # Check that column index is 3 (salary column)
            has_col_3 = bool(re.search(r',3[,)]', formula_upper))
            if has_a_to_c_range and has_col_3:
                print(f"PASS: Component 3 — Formula references A:C range with column 3 (0.3 pts)")
                total_score += 0.3
            else:
                if not has_a_to_c_range:
                    print(f"FAIL: Component 3 — Formula does not reference A:C range: {e2_val}")
                if not has_col_3:
                    print(f"FAIL: Component 3 — Formula does not retrieve column 3: {e2_val}")
        else:
            print(f"FAIL: Component 3 — E2 is not a formula string: {repr(e2_val)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'

# Persist any unsaved LibreOffice state
persist_app_state("libreoffice_calc")

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
