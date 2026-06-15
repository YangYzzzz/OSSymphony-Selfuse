"""
Initial Setup: Fleet Vehicle Maintenance Tracker
Task ID: calc_wf_029
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date, timedelta
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_029'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    wb = openpyxl.Workbook()

    # --- Style definitions ---
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # === Sheet 1: Registry ===
    ws_reg = wb.active
    ws_reg.title = "Registry"

    reg_headers = ["Vehicle ID", "Make", "Model", "Year", "Last Service Date", "Service Interval (months)"]
    for col, h in enumerate(reg_headers, 1):
        cell = ws_reg.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 10 vehicles with realistic data
    vehicles = [
        ["VH-001", "Toyota", "Camry", 2021, date(2025, 8, 15), 6],
        ["VH-002", "Ford", "F-150", 2020, date(2025, 6, 20), 4],
        ["VH-003", "Honda", "Civic", 2022, date(2025, 10, 5), 6],
        ["VH-004", "Chevrolet", "Silverado", 2019, date(2025, 3, 10), 3],
        ["VH-005", "Nissan", "Altima", 2023, date(2025, 11, 1), 6],
        ["VH-006", "Ram", "1500", 2020, date(2025, 5, 18), 4],
        ["VH-007", "Toyota", "Tacoma", 2021, date(2025, 9, 25), 6],
        ["VH-008", "Ford", "Transit", 2018, date(2025, 2, 14), 3],
        ["VH-009", "Chevrolet", "Express", 2019, date(2025, 7, 30), 4],
        ["VH-010", "Honda", "CR-V", 2022, date(2025, 12, 10), 6],
    ]

    for r, row_data in enumerate(vehicles, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_reg.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 5:
                cell.number_format = 'yyyy-mm-dd'
            if c == 4 or c == 6:
                cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws_reg.column_dimensions["A"].width = 12
    ws_reg.column_dimensions["B"].width = 14
    ws_reg.column_dimensions["C"].width = 14
    ws_reg.column_dimensions["D"].width = 8
    ws_reg.column_dimensions["E"].width = 18
    ws_reg.column_dimensions["F"].width = 22

    # === Sheet 2: Maintenance Log ===
    ws_log = wb.create_sheet("Maintenance Log")

    log_headers = ["Date", "Vehicle ID", "Service Type", "Cost", "Mileage"]
    for col, h in enumerate(log_headers, 1):
        cell = ws_log.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    service_types = [
        "Oil Change", "Tire Rotation", "Brake Inspection", "Air Filter Replacement",
        "Transmission Service", "Coolant Flush", "Battery Replacement", "Alignment",
        "Spark Plug Replacement", "Belt Replacement", "Fluid Top-Off", "Full Inspection"
    ]

    # 40 maintenance log entries spread across vehicles
    log_entries = [
        [date(2024, 1, 12), "VH-001", "Oil Change", 89.50, 32150],
        [date(2024, 2, 5), "VH-002", "Tire Rotation", 65.00, 45200],
        [date(2024, 2, 18), "VH-003", "Brake Inspection", 245.00, 18900],
        [date(2024, 3, 1), "VH-004", "Transmission Service", 420.00, 67800],
        [date(2024, 3, 15), "VH-005", "Oil Change", 92.00, 12500],
        [date(2024, 3, 22), "VH-006", "Coolant Flush", 175.00, 52300],
        [date(2024, 4, 8), "VH-007", "Air Filter Replacement", 45.00, 28700],
        [date(2024, 4, 14), "VH-008", "Battery Replacement", 289.00, 89100],
        [date(2024, 4, 28), "VH-009", "Oil Change", 95.00, 71200],
        [date(2024, 5, 10), "VH-010", "Alignment", 135.00, 22400],
        [date(2024, 5, 20), "VH-001", "Brake Inspection", 310.00, 33800],
        [date(2024, 6, 3), "VH-002", "Oil Change", 89.50, 47600],
        [date(2024, 6, 15), "VH-003", "Tire Rotation", 65.00, 20500],
        [date(2024, 6, 28), "VH-004", "Belt Replacement", 385.00, 69200],
        [date(2024, 7, 5), "VH-005", "Brake Inspection", 195.00, 14200],
        [date(2024, 7, 18), "VH-006", "Oil Change", 89.50, 54100],
        [date(2024, 7, 25), "VH-007", "Full Inspection", 350.00, 30400],
        [date(2024, 8, 2), "VH-008", "Spark Plug Replacement", 220.00, 91500],
        [date(2024, 8, 15), "VH-001", "Tire Rotation", 65.00, 35200],
        [date(2024, 8, 28), "VH-009", "Transmission Service", 445.00, 73600],
        [date(2024, 9, 5), "VH-002", "Air Filter Replacement", 48.00, 49800],
        [date(2024, 9, 18), "VH-010", "Oil Change", 92.00, 24100],
        [date(2024, 9, 25), "VH-004", "Oil Change", 95.00, 70800],
        [date(2024, 10, 3), "VH-003", "Coolant Flush", 175.00, 22100],
        [date(2024, 10, 15), "VH-006", "Tire Rotation", 65.00, 55900],
        [date(2024, 10, 22), "VH-005", "Fluid Top-Off", 35.00, 15800],
        [date(2024, 11, 1), "VH-007", "Oil Change", 89.50, 32100],
        [date(2024, 11, 12), "VH-008", "Brake Inspection", 275.00, 93200],
        [date(2024, 11, 25), "VH-009", "Air Filter Replacement", 48.00, 75100],
        [date(2024, 12, 3), "VH-001", "Full Inspection", 380.00, 36900],
        [date(2024, 12, 15), "VH-010", "Tire Rotation", 65.00, 25800],
        [date(2025, 1, 8), "VH-002", "Brake Inspection", 310.00, 51200],
        [date(2025, 1, 20), "VH-004", "Coolant Flush", 175.00, 72400],
        [date(2025, 2, 5), "VH-006", "Battery Replacement", 295.00, 57600],
        [date(2025, 2, 18), "VH-003", "Oil Change", 92.00, 23700],
        [date(2025, 3, 1), "VH-008", "Oil Change", 95.00, 94800],
        [date(2025, 3, 15), "VH-005", "Alignment", 140.00, 17200],
        [date(2025, 4, 2), "VH-009", "Belt Replacement", 365.00, 76800],
        [date(2025, 4, 18), "VH-007", "Brake Inspection", 255.00, 33800],
        [date(2025, 5, 5), "VH-010", "Spark Plug Replacement", 210.00, 27500],
    ]

    for r, row_data in enumerate(log_entries, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_log.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 1:
                cell.number_format = 'yyyy-mm-dd'
            elif c == 4:
                cell.number_format = '$#,##0.00'
            elif c == 5:
                cell.number_format = '#,##0'
            elif c == 2:
                cell.alignment = Alignment(horizontal="center")

    ws_log.column_dimensions["A"].width = 14
    ws_log.column_dimensions["B"].width = 12
    ws_log.column_dimensions["C"].width = 26
    ws_log.column_dimensions["D"].width = 12
    ws_log.column_dimensions["E"].width = 12

    # === Sheet 3: Dashboard (empty - task is to build it) ===
    ws_dash = wb.create_sheet("Dashboard")
    # Only add a title so the sheet isn't completely blank
    title_cell = ws_dash.cell(row=1, column=1, value="Fleet Maintenance Dashboard")
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="2F5496")
    ws_dash.merge_cells("A1:G1")
    title_cell.alignment = Alignment(horizontal="center")

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
