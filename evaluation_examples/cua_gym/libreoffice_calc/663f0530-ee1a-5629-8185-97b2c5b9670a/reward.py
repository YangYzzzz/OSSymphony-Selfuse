"""
Reward Script: Fleet Vehicle Maintenance Tracker
Task ID: calc_wf_029
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Dashboard headers row with correct column names
  Component 2 (0.25): 10 data rows with EDATE formulas in Next Service Due column
  Component 3 (0.15): SUMIFS formulas in Total Maintenance Cost column
  Component 4 (0.15): Overdue check formulas (IF/TODAY) in Overdue column
  Component 5 (0.10): Conditional formatting with red fill for overdue rows
  Component 6 (0.10): Bar chart present on Dashboard
"""

import os
import openpyxl
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_029'


def persist_app_state(domain):
    """Try to save any unsaved LibreOffice state."""
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print(f"PERSIST: ctrl+s sent for {domain}")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Dashboard sheet must exist
    if 'Dashboard' not in wb.sheetnames:
        print("FAIL: 'Dashboard' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Dashboard']

    # -------------------------------------------------------------------------
    # Component 1: Dashboard headers with correct column names (0.25 points)
    # The golden file has headers in row 3: Vehicle ID, Make, Model,
    # Last Service Date, Service Interval (months), Next Service Due,
    # Overdue, Total Maintenance Cost.
    # The key task-introduced headers are F (Next Service Due), G (Overdue),
    # H (Total Maintenance Cost). We check for a header row containing these.
    # -------------------------------------------------------------------------
    try:
        # Find a header row that contains the task-introduced column names
        header_row = None
        required_headers = {'next service due', 'overdue', 'total maintenance cost'}
        for r in range(1, min(ws.max_row + 1, 10)):
            row_vals = []
            for c in range(1, min(ws.max_column + 1, 12)):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    row_vals.append(str(v).strip().lower())
            found = sum(1 for h in required_headers if any(h in rv for rv in row_vals))
            if found >= 3:
                header_row = r
                break

        if header_row is not None:
            print(f"PASS: Component 1 -- Dashboard headers found in row {header_row} with required columns (0.25 pts)")
            total_score += 0.25
        else:
            print("FAIL: Component 1 -- Dashboard missing required headers (Next Service Due, Overdue, Total Maintenance Cost)")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # -------------------------------------------------------------------------
    # Component 2: 10 data rows with EDATE formulas for Next Service Due (0.25 pts)
    # The golden file has formulas like =EDATE(D4,E4) in column F rows 4-13.
    # We look for at least 8 rows with EDATE formulas in the dashboard area.
    # -------------------------------------------------------------------------
    try:
        edate_count = 0
        # Scan rows after the header for EDATE formulas
        start_row = (header_row + 1) if header_row else 2
        for r in range(start_row, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if isinstance(val, str) and 'EDATE' in val.upper():
                    edate_count += 1
                    break  # count each row once

        if edate_count >= 8:
            print(f"PASS: Component 2 -- Found {edate_count} rows with EDATE formulas (0.25 pts)")
            total_score += 0.25
        elif edate_count >= 5:
            partial = round(0.25 * edate_count / 10, 2)
            print(f"PARTIAL: Component 2 -- Found {edate_count}/10 rows with EDATE formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- Found only {edate_count} rows with EDATE formulas, need >= 8")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # -------------------------------------------------------------------------
    # Component 3: SUMIFS formulas for Total Maintenance Cost (0.15 pts)
    # The golden file uses SUMIFS on Maintenance Log to compute cost per vehicle.
    # We look for SUMIFS (or SUMIF) formulas in the dashboard data rows.
    # -------------------------------------------------------------------------
    try:
        sumifs_count = 0
        start_row = (header_row + 1) if header_row else 2
        for r in range(start_row, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if isinstance(val, str) and ('SUMIFS' in val.upper() or 'SUMIF' in val.upper()):
                    sumifs_count += 1
                    break  # count each row once

        if sumifs_count >= 8:
            print(f"PASS: Component 3 -- Found {sumifs_count} rows with SUMIFS formulas (0.15 pts)")
            total_score += 0.15
        elif sumifs_count >= 5:
            partial = round(0.15 * sumifs_count / 10, 2)
            print(f"PARTIAL: Component 3 -- Found {sumifs_count}/10 SUMIFS formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 -- Found only {sumifs_count} rows with SUMIFS formulas, need >= 8")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # -------------------------------------------------------------------------
    # Component 4: Overdue check formulas with IF/TODAY (0.15 pts)
    # The golden file uses =IF(F4<TODAY(),"OVERDUE","OK") in the Overdue column.
    # We look for IF formulas that reference TODAY in the dashboard data rows.
    # -------------------------------------------------------------------------
    try:
        overdue_formula_count = 0
        start_row = (header_row + 1) if header_row else 2
        for r in range(start_row, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if isinstance(val, str):
                    upper_val = val.upper()
                    if 'IF(' in upper_val and 'TODAY()' in upper_val:
                        overdue_formula_count += 1
                        break  # count each row once

        if overdue_formula_count >= 8:
            print(f"PASS: Component 4 -- Found {overdue_formula_count} rows with IF/TODAY overdue formulas (0.15 pts)")
            total_score += 0.15
        elif overdue_formula_count >= 5:
            partial = round(0.15 * overdue_formula_count / 10, 2)
            print(f"PARTIAL: Component 4 -- Found {overdue_formula_count}/10 IF/TODAY formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 -- Found only {overdue_formula_count} rows with IF/TODAY formulas, need >= 8")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # -------------------------------------------------------------------------
    # Component 5: Conditional formatting with red fill for overdue rows (0.10 pts)
    # The golden file has conditional formatting on the data range with a formula
    # checking for "OVERDUE" and applying red fill (FFFF0000).
    # -------------------------------------------------------------------------
    try:
        cf_rules = list(ws.conditional_formatting)
        has_overdue_cf = False
        for cf in cf_rules:
            for rule in cf.rules:
                # Check if the rule references OVERDUE or the overdue column
                formula_str = str(rule.formula) if rule.formula else ''
                rule_type = rule.type if rule.type else ''
                if 'OVERDUE' in formula_str.upper() or (
                    rule_type == 'expression' and 'OVERDUE' in formula_str.upper()
                ):
                    has_overdue_cf = True
                    break
                # Also accept any conditional formatting with red fill on the data area
                if rule.dxf and rule.dxf.fill:
                    try:
                        fill_color = rule.dxf.fill.fgColor.rgb
                        if fill_color and 'FF0000' in fill_color:
                            has_overdue_cf = True
                            break
                    except:
                        pass
            if has_overdue_cf:
                break

        if has_overdue_cf:
            print(f"PASS: Component 5 -- Conditional formatting for overdue rows found (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 -- No conditional formatting for overdue rows detected")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    # -------------------------------------------------------------------------
    # Component 6: Bar chart on Dashboard (0.10 pts)
    # The golden file has a BarChart showing maintenance cost by vehicle.
    # -------------------------------------------------------------------------
    try:
        charts = ws._charts
        if len(charts) >= 1:
            # Check that at least one chart is a bar/column chart
            has_bar = False
            for chart in charts:
                chart_class = chart.__class__.__name__
                if 'Bar' in chart_class:
                    has_bar = True
                    break
            if has_bar:
                print(f"PASS: Component 6 -- Bar chart found on Dashboard (0.10 pts)")
                total_score += 0.10
            else:
                # Accept any chart type as partial credit
                print(f"PARTIAL: Component 6 -- Chart found but not a bar chart ({chart_class}), awarding half (0.05 pts)")
                total_score += 0.05
        else:
            print(f"FAIL: Component 6 -- No charts found on Dashboard")
    except Exception as e:
        print(f"ERROR: Component 6 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
