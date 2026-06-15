"""
Initial Setup: Product return and refund processing tracker
Task ID: calc_grs_083
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_083'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Returns Log ---
    ws = wb.active
    ws.title = "Returns Log"

    headers = [
        "Return ID", "Original Order ID", "Customer Name", "Return Date",
        "Product(s) Returned", "Return Reason", "Condition", "Refund Amount",
        "Refund Method", "Processing Status", "Days to Resolution", "Notes"
    ]

    # Header styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Return data - realistic clothing e-commerce returns
    returns_data = [
        ["RTN-2025-001", "ORD-48291", "Sarah Chen", "2025-01-05", "Women's Cashmere Sweater (M)", "Wrong Item", "New/Sealed", 129.99, "Original Payment", "Closed", 3, "Received men's size instead"],
        ["RTN-2025-002", "ORD-48305", "Marcus Johnson", "2025-01-08", "Slim Fit Chinos (32x30)", "Changed Mind", "Like New", 64.50, "Store Credit", "Closed", 5, "Customer preferred different color"],
        ["RTN-2025-003", "ORD-48312", "Emily Rodriguez", "2025-01-10", "Silk Blouse (S), Linen Pants (4)", "Defective", "Damaged", 185.00, "Original Payment", "Closed", 7, "Blouse had torn seam, pants had stain"],
        ["RTN-2025-004", "ORD-48320", "James Park", "2025-01-12", "Leather Jacket (L)", "Damaged in Shipping", "Damaged", 299.00, "Original Payment", "Closed", 4, "Box was crushed, zipper broken"],
        ["RTN-2025-005", "ORD-48334", "Aisha Williams", "2025-01-15", "Cotton T-Shirt Pack (3x, M)", "Wrong Item", "New/Sealed", 45.00, "Exchange", "Closed", 2, "Received wrong color assortment"],
        ["RTN-2025-006", "ORD-48341", "David Thompson", "2025-01-18", "Wool Overcoat (XL)", "Changed Mind", "Like New", 349.00, "Store Credit", "Closed", 6, "Too formal for intended use"],
        ["RTN-2025-007", "ORD-48356", "Maria Santos", "2025-01-20", "Denim Jacket (M)", "Defective", "Used", 89.99, "Original Payment", "Disputed", 14, "Customer claims color faded after one wash"],
        ["RTN-2025-008", "ORD-48367", "Robert Kim", "2025-01-22", "Running Shoes (10.5)", "Wrong Item", "New/Sealed", 125.00, "Exchange", "Closed", 3, "Received size 9.5 instead"],
        ["RTN-2025-009", "ORD-48378", "Lisa Patel", "2025-01-25", "Evening Dress (6)", "Damaged in Shipping", "Damaged", 210.00, "Original Payment", "Closed", 5, "Dress arrived with water damage"],
        ["RTN-2025-010", "ORD-48390", "Thomas Brown", "2025-01-28", "Polo Shirt (L), Belt (34)", "Changed Mind", "New/Sealed", 78.50, "Store Credit", "Closed", 4, "Ordered duplicates by mistake"],
        ["RTN-2025-011", "ORD-48402", "Jennifer Liu", "2025-02-01", "Yoga Pants (S)", "Defective", "Used", 55.00, "Original Payment", "Closed", 8, "Seams splitting after normal wear"],
        ["RTN-2025-012", "ORD-48415", "Michael O'Brien", "2025-02-03", "Down Vest (M)", "Other", "Like New", 140.00, "Original Payment", "Closed", 3, "Allergic reaction to fill material"],
        ["RTN-2025-013", "ORD-48423", "Priya Sharma", "2025-02-06", "Maxi Skirt (M), Cardigan (M)", "Wrong Item", "New/Sealed", 112.00, "Exchange", "Closed", 4, "Both items wrong size"],
        ["RTN-2025-014", "ORD-48440", "Christopher Davis", "2025-02-10", "Flannel Shirt (XL)", "Damaged in Shipping", "Damaged", 49.99, "Original Payment", "Approved", 2, "Package was opened and resealed"],
        ["RTN-2025-015", "ORD-48451", "Fatima Al-Hassan", "2025-02-12", "Trench Coat (S)", "Defective", "Used", 275.00, "Original Payment", "Disputed", 18, "Lining detached, claims manufacturing defect"],
        ["RTN-2025-016", "ORD-48467", "Andrew Wilson", "2025-02-15", "Board Shorts (L), Rash Guard (L)", "Changed Mind", "New/Sealed", 68.00, "Store Credit", "Closed", 3, "Season changed, no longer needed"],
        ["RTN-2025-017", "ORD-48478", "Sophia Martinez", "2025-02-18", "Cashmere Scarf", "Other", "Like New", 95.00, "Exchange", "Closed", 5, "Gift recipient already had same item"],
        ["RTN-2025-018", "ORD-48490", "Daniel Lee", "2025-02-20", "Dress Shirt (16/34)", "Wrong Item", "New/Sealed", 79.99, "Original Payment", "Initiated", 1, "Received casual shirt instead"],
        ["RTN-2025-019", "ORD-48503", "Rachel Green", "2025-02-22", "Ankle Boots (8)", "Defective", "Used", 165.00, "Original Payment", "Refund Issued", 10, "Sole detaching after two weeks"],
        ["RTN-2025-020", "ORD-48515", "Kevin Nakamura", "2025-02-25", "Swim Trunks (M)", "Damaged in Shipping", "Damaged", 35.00, "Exchange", "Approved", 2, "Item arrived wet and stained"],
        ["RTN-2025-021", "ORD-48530", "Amanda Foster", "2025-03-01", "Blazer (10), Slacks (10)", "Changed Mind", "New/Sealed", 245.00, "Store Credit", "Initiated", 1, "Found better deal elsewhere"],
        ["RTN-2025-022", "ORD-48542", "Hassan Ahmed", "2025-03-04", "Henley Shirt (M)", "Defective", "Used", 42.00, "Original Payment", "Disputed", 12, "Claims shrinkage after washing per instructions"],
        ["RTN-2025-023", "ORD-48555", "Olivia Taylor", "2025-03-07", "Wrap Dress (8)", "Wrong Item", "New/Sealed", 155.00, "Original Payment", "Refund Issued", 6, "Color completely different from website photo"],
        ["RTN-2025-024", "ORD-48568", "Brian Cooper", "2025-03-10", "Cargo Pants (34x32)", "Other", "Like New", 59.99, "Store Credit", "Closed", 4, "Pockets too small for stated utility design"],
        ["RTN-2025-025", "ORD-48580", "Nina Volkov", "2025-03-12", "Puffer Jacket (XS)", "Damaged in Shipping", "Damaged", 189.00, "Original Payment", "Approved", 3, "Jacket torn, feathers leaking out"],
    ]

    for r, row_data in enumerate(returns_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 4:  # Return Date
                cell.number_format = 'yyyy-mm-dd'
            elif c == 8:  # Refund Amount
                cell.number_format = '$#,##0.00'
            elif c == 11:  # Days to Resolution - raw numbers, no formula
                cell.number_format = '0'

    # Set column widths
    col_widths = {
        'A': 16, 'B': 16, 'C': 20, 'D': 14, 'E': 35, 'F': 22,
        'G': 16, 'H': 14, 'I': 20, 'J': 18, 'K': 18, 'L': 40
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Row height for header
    ws.row_dimensions[1].height = 30

    # Freeze header row
    ws.freeze_panes = "A2"

    # Data validations (dropdowns)
    dv_reason = DataValidation(
        type="list",
        formula1='"Defective,Wrong Item,Changed Mind,Damaged in Shipping,Other"',
        allow_blank=True, showDropDown=False
    )
    dv_reason.prompt = "Select return reason"
    dv_reason.promptTitle = "Return Reason"
    dv_reason.add("F2:F100")
    ws.add_data_validation(dv_reason)

    dv_condition = DataValidation(
        type="list",
        formula1='"New/Sealed,Like New,Used,Damaged"',
        allow_blank=True, showDropDown=False
    )
    dv_condition.prompt = "Select item condition"
    dv_condition.promptTitle = "Condition"
    dv_condition.add("G2:G100")
    ws.add_data_validation(dv_condition)

    dv_refund_method = DataValidation(
        type="list",
        formula1='"Original Payment,Store Credit,Exchange"',
        allow_blank=True, showDropDown=False
    )
    dv_refund_method.prompt = "Select refund method"
    dv_refund_method.promptTitle = "Refund Method"
    dv_refund_method.add("I2:I100")
    ws.add_data_validation(dv_refund_method)

    dv_status = DataValidation(
        type="list",
        formula1='"Initiated,Approved,Refund Issued,Closed,Disputed"',
        allow_blank=True, showDropDown=False
    )
    dv_status.prompt = "Select processing status"
    dv_status.promptTitle = "Processing Status"
    dv_status.add("J2:J100")
    ws.add_data_validation(dv_status)

    # --- Sheet 2: Summary (blank placeholder) ---
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Summary"
    ws2["A1"].font = Font(name="Calibri", size=14, bold=True)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
