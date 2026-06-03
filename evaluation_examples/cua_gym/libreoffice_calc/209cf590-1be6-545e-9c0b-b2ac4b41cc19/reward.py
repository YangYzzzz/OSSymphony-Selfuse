"""
Reward Script: Product Return and Refund Processing Tracker
Task ID: calc_grs_083
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): K column "Days to Resolution" contains formulas (not plain numbers)
  Component 2 (0.20): Conditional formatting highlights Disputed rows in red
  Component 3 (0.20): Summary sheet has Return Rate by Reason section with COUNTIF formulas
  Component 4 (0.15): Summary sheet has Average Resolution Time with AVERAGE/MIN/MAX
  Component 5 (0.10): Summary sheet has Refund Amounts by Month with SUMPRODUCT formulas
  Component 6 (0.10): Summary sheet has a chart for return reasons distribution
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_083'


def persist_app_state(domain):
    """Save any unsaved LibreOffice state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print(f"PERSIST: ctrl+s sent for {domain}")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Returns Log sheet must exist
    if 'Returns Log' not in wb.sheetnames:
        print("CRITICAL: 'Returns Log' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Returns Log']

    # ================================================================
    # Component 1: K column "Days to Resolution" has formulas (0.25 pts)
    # In initial_env, K2:K26 contain plain integers.
    # In golden_env, they contain IF formulas for dynamic calculation.
    # We check that at least 15 of the K cells (K2:K26) contain formula strings.
    # ================================================================
    try:
        formula_count = 0
        total_k_cells = 0
        for row_num in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row_num, column=11).value  # K column
            if cell_val is not None:
                total_k_cells += 1
                if isinstance(cell_val, str) and cell_val.startswith('='):
                    formula_count += 1

        if total_k_cells > 0 and formula_count >= 15:
            print(f"PASS: Component 1 — K column has {formula_count}/{total_k_cells} formula cells (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — K column has {formula_count}/{total_k_cells} formula cells, need >= 15")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ================================================================
    # Component 2: Conditional formatting for Disputed rows in red (0.20 pts)
    # In initial_env, there is NO conditional formatting.
    # In golden_env, there is a formula-based CF rule that checks for "Disputed"
    # in column J and applies red fill.
    # ================================================================
    try:
        cf_list = list(ws.conditional_formatting)
        found_dispute_cf = False

        for cf in cf_list:
            for rule in cf.rules:
                # Check if the rule references "Disputed" in its formula
                rule_formulas = rule.formula if rule.formula else []
                for f in rule_formulas:
                    if 'Disputed' in str(f) or 'disputed' in str(f).lower():
                        # Check that it has a red fill in the differential style
                        if rule.dxf and rule.dxf.fill:
                            fg = rule.dxf.fill.fgColor
                            if fg and fg.rgb:
                                rgb_str = str(fg.rgb).upper()
                                # FFFF0000 = opaque red
                                if 'FF0000' in rgb_str:
                                    found_dispute_cf = True

        if found_dispute_cf:
            print(f"PASS: Component 2 — Conditional formatting for Disputed rows with red fill found (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 — No conditional formatting rule found that highlights Disputed rows in red")
            # Partial: check if any CF rule references Disputed at all
            for cf in cf_list:
                for rule in cf.rules:
                    rule_formulas = rule.formula if rule.formula else []
                    for f in rule_formulas:
                        if 'Disputed' in str(f) or 'disputed' in str(f).lower():
                            print(f"  (Partial: found CF rule with Disputed formula but fill is not red)")
                            total_score += 0.10
                            break
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ================================================================
    # Component 3: Summary - Return Rate by Reason with COUNTIF formulas (0.20 pts)
    # In initial_env, Summary sheet has only A1="Summary", no formulas.
    # In golden_env, Summary has COUNTIF formulas for each return reason.
    # ================================================================
    try:
        if 'Summary' not in wb.sheetnames:
            print("FAIL: Component 3 — 'Summary' sheet not found")
        else:
            ws_sum = wb['Summary']

            # Check for COUNTIF formulas referencing return reasons
            countif_reasons_found = 0
            expected_reasons = ['Defective', 'Wrong Item', 'Changed Mind', 'Damaged in Shipping', 'Other']

            for row in ws_sum.iter_rows(min_row=1, max_row=ws_sum.max_row, max_col=ws_sum.max_column, values_only=False):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        val_upper = cell.value.upper()
                        if 'COUNTIF' in val_upper:
                            for reason in expected_reasons:
                                if reason.lower() in cell.value.lower() or reason in cell.value:
                                    countif_reasons_found += 1

            # Also check for percentage formulas (division by sum)
            pct_formulas = 0
            for row in ws_sum.iter_rows(min_row=1, max_row=ws_sum.max_row, max_col=ws_sum.max_column, values_only=False):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and '/' in cell.value and 'SUM' in cell.value.upper():
                        pct_formulas += 1

            if countif_reasons_found >= 4:
                print(f"PASS: Component 3 — Found {countif_reasons_found} COUNTIF formulas for return reasons (0.20 pts)")
                total_score += 0.20
            elif countif_reasons_found >= 2:
                print(f"PARTIAL: Component 3 — Found {countif_reasons_found} COUNTIF formulas (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 3 — Found {countif_reasons_found} COUNTIF formulas for return reasons")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ================================================================
    # Component 4: Summary - Average Resolution Time section (0.15 pts)
    # Golden has AVERAGE, MIN, MAX formulas referencing Returns Log K column.
    # Initial Summary has none of these.
    # ================================================================
    try:
        if 'Summary' not in wb.sheetnames:
            print("FAIL: Component 4 — 'Summary' sheet not found")
        else:
            ws_sum = wb['Summary']
            found_avg = False
            found_min_or_max = False

            for row in ws_sum.iter_rows(min_row=1, max_row=ws_sum.max_row, max_col=ws_sum.max_column, values_only=False):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        val_upper = cell.value.upper()
                        if 'AVERAGE' in val_upper and 'K' in val_upper.replace('BLANK', ''):
                            found_avg = True
                        if ('MIN' in val_upper or 'MAX' in val_upper) and 'K' in val_upper:
                            found_min_or_max = True

            if found_avg and found_min_or_max:
                print(f"PASS: Component 4 — Found AVERAGE and MIN/MAX formulas for resolution time (0.15 pts)")
                total_score += 0.15
            elif found_avg:
                print(f"PARTIAL: Component 4 — Found AVERAGE but no MIN/MAX (0.08 pts)")
                total_score += 0.08
            else:
                print(f"FAIL: Component 4 — No AVERAGE formula for resolution time found in Summary")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ================================================================
    # Component 5: Summary - Refund Amounts by Month (0.10 pts)
    # Golden has SUMPRODUCT or SUMIFS formulas for monthly refund totals.
    # Initial Summary has none of these.
    # ================================================================
    try:
        if 'Summary' not in wb.sheetnames:
            print("FAIL: Component 5 — 'Summary' sheet not found")
        else:
            ws_sum = wb['Summary']
            monthly_formulas = 0

            for row in ws_sum.iter_rows(min_row=1, max_row=ws_sum.max_row, max_col=ws_sum.max_column, values_only=False):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        val_upper = cell.value.upper()
                        # Look for SUMPRODUCT, SUMIFS, or SUMIF referencing dates/amounts
                        if ('SUMPRODUCT' in val_upper or 'SUMIFS' in val_upper or 'SUMIF' in val_upper) and ('H' in val_upper or 'REFUND' in val_upper.replace('RETURNS LOG', '')):
                            monthly_formulas += 1

            if monthly_formulas >= 2:
                print(f"PASS: Component 5 — Found {monthly_formulas} monthly refund sum formulas (0.10 pts)")
                total_score += 0.10
            elif monthly_formulas >= 1:
                print(f"PARTIAL: Component 5 — Found {monthly_formulas} monthly refund formula (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 5 — No monthly refund aggregation formulas found in Summary")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # ================================================================
    # Component 6: Chart on Summary sheet for return reasons distribution (0.10 pts)
    # Initial has 0 charts. Golden has 1 PieChart.
    # ================================================================
    try:
        if 'Summary' not in wb.sheetnames:
            print("FAIL: Component 6 — 'Summary' sheet not found")
        else:
            ws_sum = wb['Summary']
            chart_count = len(ws_sum._charts)

            if chart_count >= 1:
                # Verify it's related to return reasons (check chart title or type)
                chart = ws_sum._charts[0]
                chart_info = str(type(chart).__name__)
                print(f"PASS: Component 6 — Found {chart_count} chart(s) on Summary ({chart_info}) (0.10 pts)")
                total_score += 0.10
            else:
                # Also check if chart is on a different sheet
                chart_found_elsewhere = False
                for sheet_name in wb.sheetnames:
                    ws_check = wb[sheet_name]
                    if sheet_name != 'Returns Log' and len(ws_check._charts) > 0:
                        chart_found_elsewhere = True
                        print(f"PARTIAL: Component 6 — Chart found on '{sheet_name}' instead of Summary (0.05 pts)")
                        total_score += 0.05
                        break
                if not chart_found_elsewhere:
                    print(f"FAIL: Component 6 — No charts found on Summary sheet (or any other non-log sheet)")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
