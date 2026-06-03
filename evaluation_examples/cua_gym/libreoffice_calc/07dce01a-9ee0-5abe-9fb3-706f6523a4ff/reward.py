"""
Reward Script: Configure 'Report Card' sheet print settings
Task ID: calc_mcp_088
Domain: libreoffice_calc
Scoring:
  - Component 1: Paper width = 8in (0.25 pts)
  - Component 2: Paper height = 11in (0.25 pts)
  - Component 3: Orientation = portrait (0.2 pts)
  - Component 4: Top margin = 0.75 inches (0.15 pts)
  - Component 5: Bottom margin = 0.75 inches (0.15 pts)
Note: left/right margins are already 0.75 in initial state, so they
are preconditions and NOT scored independently.
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_088'


def _check_dimension(raw_value, expected_inches):
    """Check if a page dimension (e.g. paperWidth) matches expected inches."""
    if raw_value is None:
        return False
    s = str(raw_value).strip().lower()
    if s == f'{int(expected_inches)}in':
        return s == f'{int(expected_inches)}in'
    try:
        num = float(s.replace('in', ''))
        return abs(num - expected_inches) < 0.1
    except (ValueError, TypeError):
        return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Report Card' sheet must exist
    if 'Report Card' not in wb.sheetnames:
        print(f"FAIL: 'Report Card' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Report Card']
    ps = ws.page_setup
    pm = ws.page_margins

    # Component 1: Paper width = 8in (0.25 points)
    try:
        paper_width = ps.paperWidth
        # openpyxl stores paperWidth as a string like "8in" or a numeric value
        if _check_dimension(paper_width, 8.0):
            print(f"PASS: Component 1 -- Paper width is 8in (found: {paper_width}) (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 -- Paper width expected 8in, found: {paper_width}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Paper height = 11in (0.25 points)
    try:
        paper_height = ps.paperHeight
        if _check_dimension(paper_height, 11.0):
            print(f"PASS: Component 2 -- Paper height is 11in (found: {paper_height}) (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 -- Paper height expected 11in, found: {paper_height}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Orientation = portrait (0.2 points)
    try:
        orientation = ps.orientation
        if orientation is not None and str(orientation).strip().lower() == 'portrait':
            print(f"PASS: Component 3 -- Orientation is portrait (found: {orientation}) (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 -- Orientation expected 'portrait', found: {orientation}")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Top margin = 0.75 inches (0.15 points)
    # Initial state has top=1.0, golden has top=0.75 -- this is a task-introduced change
    try:
        top_margin = pm.top
        if top_margin is not None and abs(float(top_margin) - 0.75) < 0.01:
            print(f"PASS: Component 4 -- Top margin is 0.75 (found: {top_margin}) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 -- Top margin expected 0.75, found: {top_margin}")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # Component 5: Bottom margin = 0.75 inches (0.15 points)
    # Initial state has bottom=1.0, golden has bottom=0.75 -- this is a task-introduced change
    try:
        bottom_margin = pm.bottom
        if bottom_margin is not None and abs(float(bottom_margin) - 0.75) < 0.01:
            print(f"PASS: Component 5 -- Bottom margin is 0.75 (found: {bottom_margin}) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 5 -- Bottom margin expected 0.75, found: {bottom_margin}")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
