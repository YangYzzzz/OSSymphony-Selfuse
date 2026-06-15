"""
Initial Setup: Configure print settings for Report Card sheet
Task ID: calc_mcp_088
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_088'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Report Card ---
    ws = wb.active
    ws.title = 'Report Card'

    # Header styling
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    # Title row (merged)
    ws.merge_cells('A1:G1')
    ws['A1'] = 'Westfield Academy - Student Report Card (Spring 2025)'
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='1F3864')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Subtitle
    ws.merge_cells('A2:G2')
    ws['A2'] = 'Grade 10 - Section B'
    ws['A2'].font = Font(name='Calibri', size=11, italic=True, color='404040')
    ws['A2'].alignment = Alignment(horizontal='center')

    # Headers in row 4
    headers = ['Student Name', 'Mathematics', 'Science', 'English', 'History', 'Art', 'Overall GPA']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Student data
    students = [
        ['Emily Rodriguez', 92, 88, 95, 87, 91, None],
        ['James Nakamura', 78, 94, 82, 90, 76, None],
        ['Priya Sharma', 96, 97, 93, 88, 85, None],
        ['Marcus Thompson', 71, 68, 85, 79, 92, None],
        ['Sofia Petrov', 88, 91, 90, 94, 89, None],
        ['Aiden O\'Brien', 65, 72, 78, 83, 95, None],
        ['Mei-Ling Chen', 99, 96, 91, 92, 88, None],
        ['Dante Williams', 84, 79, 76, 71, 82, None],
        ['Zara Al-Rashid', 90, 85, 97, 96, 93, None],
        ['Lucas Fernandez', 73, 81, 69, 77, 84, None],
        ['Hannah Kim', 95, 93, 88, 85, 79, None],
        ['Oliver Jensen', 82, 76, 91, 88, 90, None],
    ]

    data_font = Font(name='Calibri', size=11)
    center_align = Alignment(horizontal='center', vertical='center')
    name_align = Alignment(horizontal='left', vertical='center')

    for r, student in enumerate(students, 5):
        for c, val in enumerate(student, 1):
            cell = ws.cell(row=r, column=c)
            if c == 1:
                cell.value = val
                cell.alignment = name_align
            elif c == 7:
                # GPA formula: average of the 5 subjects
                cell.value = f'=ROUND(AVERAGE(B{r}:F{r}),1)'
                cell.alignment = center_align
                cell.number_format = '0.0'
            else:
                cell.value = val
                cell.alignment = center_align
            cell.font = data_font

    # Add light alternating row colors
    light_fill = PatternFill(start_color='FFD6E4F0', end_color='FFD6E4F0', fill_type='solid')
    for r in range(5, 17):
        if r % 2 == 0:
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = light_fill

    # Summary row
    summary_row = 18
    ws.cell(row=summary_row, column=1, value='Class Average').font = Font(name='Calibri', size=11, bold=True)
    for c in range(2, 8):
        col_letter = chr(64 + c)
        cell = ws.cell(row=summary_row, column=c)
        cell.value = f'=ROUND(AVERAGE({col_letter}5:{col_letter}16),1)'
        cell.font = Font(name='Calibri', size=11, bold=True)
        cell.alignment = center_align
        cell.number_format = '0.0'

    # Column widths
    ws.column_dimensions['A'].width = 22
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].width = 14

    # Row heights
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[4].height = 22

    # Add thin borders to data area
    thin_side = Side(style='thin', color='B4B4B4')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    for r in range(4, summary_row + 1):
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = thin_border

    # Leave page setup at DEFAULT (no custom paper size, no custom margins)
    # Default: Letter size, default margins (top=1.0, bottom=1.0, left=0.75, right=0.75)
    # The task requires the agent to change these settings.

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
