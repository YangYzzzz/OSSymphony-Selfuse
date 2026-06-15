"""
Initial Setup: Nested IFERROR/VLOOKUP formula across Primary and Secondary catalogs
Task ID: calc_fma_nested_iferror_063
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fma_nested_iferror_063'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Lookup ---
    ws_lookup = wb.active
    ws_lookup.title = 'Lookup'

    # Headers
    ws_lookup['A1'] = 'Product Code'
    ws_lookup['B1'] = 'Product Name'

    # Product codes PRD-001 through PRD-012 (B2:B13 are intentionally empty)
    product_codes = [
        'PRD-001',
        'PRD-002',
        'PRD-003',
        'PRD-004',
        'PRD-005',
        'PRD-006',
        'PRD-007',
        'PRD-008',
        'PRD-009',
        'PRD-010',
        'PRD-011',
        'PRD-012',
    ]
    for r, code in enumerate(product_codes, 2):
        ws_lookup.cell(row=r, column=1, value=code)
    # B2:B13 remain empty (the agent must fill them in)

    # --- Sheet 2: Primary ---
    ws_primary = wb.create_sheet('Primary')

    ws_primary['A1'] = 'Code'
    ws_primary['B1'] = 'Name'

    primary_data = [
        ('PRD-001', 'Alpha'),
        ('PRD-002', 'Beta'),
        ('PRD-003', 'Gamma'),
        ('PRD-005', 'Epsilon'),
        ('PRD-007', 'Eta'),
        ('PRD-009', 'Iota'),
    ]
    for r, (code, name) in enumerate(primary_data, 2):
        ws_primary.cell(row=r, column=1, value=code)
        ws_primary.cell(row=r, column=2, value=name)

    # --- Sheet 3: Secondary ---
    ws_secondary = wb.create_sheet('Secondary')

    ws_secondary['A1'] = 'Code'
    ws_secondary['B1'] = 'Name'

    secondary_data = [
        ('PRD-004', 'Delta'),
        ('PRD-006', 'Zeta'),
        ('PRD-008', 'Theta'),
        ('PRD-010', 'Kappa'),
        ('PRD-011', 'Lambda'),
    ]
    for r, (code, name) in enumerate(secondary_data, 2):
        ws_secondary.cell(row=r, column=1, value=code)
        ws_secondary.cell(row=r, column=2, value=name)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
