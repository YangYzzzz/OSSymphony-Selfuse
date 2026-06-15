"""
Reward Script: Nested IFERROR VLOOKUP lookup across Primary and Secondary catalogs
Task ID: calc_fma_nested_iferror_063
Domain: libreoffice_calc
Scoring:
  Component 1 (0.50): All 12 cells B2:B13 contain a nested IFERROR/VLOOKUP formula structure
  Component 2 (0.30): Formulas reference both Primary and Secondary sheets with correct column index (2)
  Component 3 (0.20): Formula in B13 includes "Discontinued" fallback (for PRD-012, not in either catalog)
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fma_nested_iferror_063'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify that the required sheets exist (precondition gate)
    required_sheets = ['Lookup', 'Primary', 'Secondary']
    for sheet in required_sheets:
        if sheet not in wb.sheetnames:
            print(f"CRITICAL: Required sheet '{sheet}' not found. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0

    ws = wb['Lookup']

    # -------------------------------------------------------------------------
    # Component 1: All 12 cells B2:B13 contain a nested IFERROR formula (0.50 points)
    # The task asks to fill B2:B13 with formulas that implement lookup with fallback.
    # Initial file has all empty cells. Golden file has formulas in all 12 cells.
    # -------------------------------------------------------------------------
    try:
        formulas_present = 0
        iferror_formulas = 0
        vlookup_formulas = 0
        missing_cells = []

        for row in range(2, 14):  # rows 2..13
            cell_val = ws.cell(row=row, column=2).value
            if cell_val is None or cell_val == '':
                missing_cells.append(f"B{row}")
                continue
            formulas_present += 1
            val_upper = str(cell_val).upper().replace(' ', '')
            if 'IFERROR' in val_upper:
                iferror_formulas += 1
            if 'VLOOKUP' in val_upper:
                vlookup_formulas += 1

        if formulas_present == 12 and iferror_formulas == 12 and vlookup_formulas >= 12:
            print(f"PASS: Component 1 — All 12 cells B2:B13 contain nested IFERROR/VLOOKUP formulas (0.50 pts)")
            total_score += 0.50
        elif formulas_present == 12:
            # Formulas present but may not use IFERROR
            print(f"FAIL: Component 1 — 12 cells filled but IFERROR count={iferror_formulas}/12, VLOOKUP count={vlookup_formulas}/12")
        else:
            print(f"FAIL: Component 1 — Only {formulas_present}/12 cells have formulas. Missing: {missing_cells}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Formulas reference both Primary and Secondary sheets with
    # correct column index (2) for the lookup (0.30 points)
    # -------------------------------------------------------------------------
    try:
        correct_refs = 0
        checked = 0

        for row in range(2, 14):
            cell_val = ws.cell(row=row, column=2).value
            if cell_val is None:
                continue
            val_str = str(cell_val)
            val_upper_nospace = val_str.upper().replace(' ', '')
            checked += 1

            # Check that formula references Primary and Secondary with col index 2 and exact match (0)
            has_primary = 'PRIMARY' in val_upper_nospace
            has_secondary = 'SECONDARY' in val_upper_nospace
            # Column index 2 means lookup returns 2nd column (name)
            # Exact match in VLOOKUP is indicated by 0 or FALSE
            has_col2 = ',2,0)' in val_upper_nospace or ',2,FALSE)' in val_upper_nospace
            # Should have both Primary and Secondary lookup
            if has_primary and has_secondary and has_col2:
                correct_refs += 1

        if checked > 0 and correct_refs == checked:
            print(f"PASS: Component 2 — All {correct_refs}/{checked} formulas reference Primary and Secondary with column index 2 (0.30 pts)")
            total_score += 0.30
        elif checked > 0:
            print(f"FAIL: Component 2 — Only {correct_refs}/{checked} formulas reference both Primary+Secondary with correct col index 2")
        else:
            print(f"FAIL: Component 2 — No formulas found to check (B2:B13 are empty)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: The "Discontinued" fallback is present in the formulas (0.20 points)
    # B13 (PRD-012) is not in either catalog, so its formula must include "Discontinued"
    # as the final fallback. This verifies the nested IFERROR structure is correct.
    # -------------------------------------------------------------------------
    try:
        b13_val = ws.cell(row=13, column=2).value
        if b13_val is not None:
            val_upper = str(b13_val).upper().replace(' ', '')
            if '"DISCONTINUED"' in val_upper or "'DISCONTINUED'" in val_upper or 'DISCONTINUED' in val_upper:
                print(f'PASS: Component 3 — B13 formula includes "Discontinued" fallback (0.20 pts)')
                total_score += 0.20
            else:
                print(f'FAIL: Component 3 — B13 formula found but does not include "Discontinued": {repr(b13_val)}')
        else:
            print(f'FAIL: Component 3 — B13 is empty, no "Discontinued" fallback found')
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
