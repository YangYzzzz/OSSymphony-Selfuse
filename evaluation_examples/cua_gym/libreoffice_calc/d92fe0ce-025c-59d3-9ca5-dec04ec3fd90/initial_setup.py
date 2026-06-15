"""
Initial Setup: Customer order form with auto-calculations
Task ID: calc_wf_019
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_019'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ========== Sheet 1: Catalog (25 products) ==========
    ws_cat = wb.active
    ws_cat.title = 'Catalog'

    # Headers
    cat_headers = ['Code', 'Name', 'Price', 'Weight']
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(cat_headers, 1):
        cell = ws_cat.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 25 products
    products = [
        ['WDG-001', 'Stainless Steel Water Bottle', 24.99, 0.8],
        ['WDG-002', 'Bamboo Cutting Board Set', 34.50, 2.5],
        ['WDG-003', 'Ceramic Coffee Mug (16oz)', 12.99, 0.6],
        ['WDG-004', 'Cast Iron Skillet 12"', 49.95, 3.8],
        ['WDG-005', 'Electric Kettle 1.7L', 39.99, 1.2],
        ['ELC-001', 'Wireless Bluetooth Earbuds', 59.99, 0.2],
        ['ELC-002', 'USB-C Charging Hub (7-port)', 45.00, 0.4],
        ['ELC-003', 'LED Desk Lamp with Dimmer', 32.50, 1.5],
        ['ELC-004', 'Portable Power Bank 20000mAh', 29.99, 0.5],
        ['ELC-005', 'Smart WiFi Plug (3-pack)', 27.99, 0.3],
        ['OFF-001', 'Ergonomic Gel Mouse Pad', 15.99, 0.3],
        ['OFF-002', 'Adjustable Monitor Stand', 42.00, 2.8],
        ['OFF-003', 'Premium Ballpoint Pen Set', 18.50, 0.2],
        ['OFF-004', 'Desktop File Organizer', 28.99, 1.8],
        ['OFF-005', 'Noise-Canceling Headphones', 89.99, 0.4],
        ['HOM-001', 'Scented Soy Candle Set (3)', 22.99, 1.2],
        ['HOM-002', 'Microfiber Throw Blanket', 35.00, 1.0],
        ['HOM-003', 'Decorative Wall Clock', 27.50, 0.9],
        ['HOM-004', 'Indoor Herb Garden Kit', 31.99, 2.0],
        ['HOM-005', 'Aromatherapy Diffuser', 38.00, 0.7],
        ['SPT-001', 'Yoga Mat with Carry Strap', 29.99, 1.5],
        ['SPT-002', 'Resistance Band Set (5)', 19.99, 0.6],
        ['SPT-003', 'Insulated Sports Bottle 32oz', 22.50, 0.5],
        ['SPT-004', 'Foam Roller 18"', 24.99, 0.8],
        ['SPT-005', 'Digital Jump Rope', 16.99, 0.3],
    ]

    for r, row_data in enumerate(products, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_cat.cell(row=r, column=c, value=val)
            if c == 3:  # Price column
                cell.number_format = '$#,##0.00'
            elif c == 4:  # Weight column
                cell.number_format = '0.0'

    ws_cat.column_dimensions['A'].width = 12
    ws_cat.column_dimensions['B'].width = 35
    ws_cat.column_dimensions['C'].width = 12
    ws_cat.column_dimensions['D'].width = 10

    # ========== Sheet 2: Order Form ==========
    ws_order = wb.create_sheet('Order Form')

    # --- Company Header (Rows 1-3) ---
    ws_order.merge_cells('A1:E1')
    ws_order['A1'] = 'Summit Gear Supply Co.'
    ws_order['A1'].font = Font(name='Arial', size=18, bold=True, color='2F5496')
    ws_order['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws_order.merge_cells('A2:E2')
    ws_order['A2'] = '742 Mountain View Drive, Boulder, CO 80301'
    ws_order['A2'].font = Font(name='Arial', size=10, color='666666')
    ws_order['A2'].alignment = Alignment(horizontal='center')

    ws_order.merge_cells('A3:E3')
    ws_order['A3'] = 'Tel: (303) 555-0187  |  orders@summitgear.com  |  www.summitgear.com'
    ws_order['A3'].font = Font(name='Arial', size=9, color='666666')
    ws_order['A3'].alignment = Alignment(horizontal='center')

    # Row 4: separator (blank)

    # --- Customer Info (Rows 5-7) ---
    label_font = Font(name='Arial', size=10, bold=True)
    info_font = Font(name='Arial', size=10)

    ws_order['A5'] = 'Customer Name:'
    ws_order['A5'].font = label_font
    ws_order['B5'] = 'Rebecca Martinez'
    ws_order['B5'].font = info_font

    ws_order['D5'] = 'Order Date:'
    ws_order['D5'].font = label_font
    ws_order['E5'] = '2025-11-15'
    ws_order['E5'].font = info_font

    ws_order['A6'] = 'Address:'
    ws_order['A6'].font = label_font
    ws_order['B6'] = '1580 Elm Street, Denver, CO 80202'
    ws_order['B6'].font = info_font

    ws_order['D6'] = 'Order #:'
    ws_order['D6'].font = label_font
    ws_order['E6'] = 'ORD-2025-4782'
    ws_order['E6'].font = info_font

    ws_order['A7'] = 'Phone:'
    ws_order['A7'].font = label_font
    ws_order['B7'] = '(720) 555-0293'
    ws_order['B7'].font = info_font

    ws_order['D7'] = 'Payment:'
    ws_order['D7'].font = label_font
    ws_order['E7'] = 'Net 30'
    ws_order['E7'].font = info_font

    # Row 8: blank separator

    # --- Order Table Header (Row 9) ---
    order_headers = ['Code', 'Product', 'Price', 'Qty', 'Line Total']
    order_header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    order_header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(order_headers, 1):
        cell = ws_order.cell(row=9, column=col, value=h)
        cell.font = order_header_font
        cell.fill = order_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # --- Order Lines (Rows 10-20) ---
    # Pre-fill some product codes for realism, but NO formulas
    sample_codes = [
        'WDG-001', 'ELC-002', 'OFF-005', 'HOM-003', 'SPT-001',
        'WDG-004', 'ELC-001', '', '', '', ''
    ]
    sample_qty = [2, 1, 1, 3, 2, 1, 2, None, None, None, None]

    for i, (code, qty) in enumerate(zip(sample_codes, sample_qty)):
        row = 10 + i
        # Code column
        cell_code = ws_order.cell(row=row, column=1, value=code if code else None)
        cell_code.border = thin_border
        cell_code.alignment = Alignment(horizontal='center')

        # Product column (empty - task needs VLOOKUP here)
        cell_prod = ws_order.cell(row=row, column=2)
        cell_prod.border = thin_border

        # Price column (empty - task needs VLOOKUP here)
        cell_price = ws_order.cell(row=row, column=3)
        cell_price.border = thin_border
        cell_price.number_format = '$#,##0.00'

        # Qty column
        cell_qty = ws_order.cell(row=row, column=4, value=qty)
        cell_qty.border = thin_border
        cell_qty.alignment = Alignment(horizontal='center')

        # Line Total column (empty - task needs formula here)
        cell_total = ws_order.cell(row=row, column=5)
        cell_total.border = thin_border
        cell_total.number_format = '$#,##0.00'

    # --- Bottom Section (after row 20) ---
    # Row 22: Subtotal
    ws_order['D22'] = 'Subtotal:'
    ws_order['D22'].font = Font(name='Arial', size=11, bold=True)
    ws_order['D22'].alignment = Alignment(horizontal='right')
    ws_order['E22'].border = thin_border
    ws_order['E22'].number_format = '$#,##0.00'

    # Row 23: Discount
    ws_order['D23'] = 'Discount:'
    ws_order['D23'].font = Font(name='Arial', size=11, bold=True)
    ws_order['D23'].alignment = Alignment(horizontal='right')
    ws_order['E23'].border = thin_border
    ws_order['E23'].number_format = '$#,##0.00'

    # Row 24: Shipping
    ws_order['D24'] = 'Shipping:'
    ws_order['D24'].font = Font(name='Arial', size=11, bold=True)
    ws_order['D24'].alignment = Alignment(horizontal='right')
    ws_order['E24'].border = thin_border
    ws_order['E24'].number_format = '$#,##0.00'

    # Row 25: Grand Total
    ws_order['D25'] = 'Grand Total:'
    ws_order['D25'].font = Font(name='Arial', size=12, bold=True, color='2F5496')
    ws_order['D25'].alignment = Alignment(horizontal='right')
    ws_order['E25'].border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='double', color='2F5496'),
        bottom=Side(style='double', color='2F5496'),
    )
    ws_order['E25'].number_format = '$#,##0.00'
    ws_order['E25'].font = Font(name='Arial', size=12, bold=True, color='2F5496')

    # --- Column widths ---
    ws_order.column_dimensions['A'].width = 14
    ws_order.column_dimensions['B'].width = 35
    ws_order.column_dimensions['C'].width = 14
    ws_order.column_dimensions['D'].width = 16
    ws_order.column_dimensions['E'].width = 16

    # Row heights for header area
    ws_order.row_dimensions[1].height = 30

    # Catalog sheet is VISIBLE in initial (task asks to hide it)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
