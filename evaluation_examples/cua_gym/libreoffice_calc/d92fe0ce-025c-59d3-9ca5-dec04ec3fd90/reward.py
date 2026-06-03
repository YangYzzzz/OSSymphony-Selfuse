"""
Reward Script: Customer order form with auto-calculations
Task ID: calc_wf_019
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): VLOOKUP formulas in B and C columns for product lookup
  Component 2 (0.20): Line total formulas in E column (price * qty)
  Component 3 (0.25): Summary formulas (Subtotal, Discount, Shipping, Grand Total)
  Component 4 (0.15): Catalog sheet hidden
  Component 5 (0.15): Print area set on Order Form
"""

import os

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_019'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: both sheets must exist
    if 'Catalog' not in wb.sheetnames or 'Order Form' not in wb.sheetnames:
        print(f"FAIL: Required sheets missing. Found: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Order Form']

    # ---------------------------------------------------------------
    # Component 1: VLOOKUP formulas in B and C columns (0.25 points)
    # Initial: B10:B16 and C10:C16 are None (empty)
    # Golden: VLOOKUP formulas to retrieve product name and price
    # ---------------------------------------------------------------
    try:
        vlookup_count = 0
        total_lookup_cells = 0
        for row in range(10, 17):  # rows 10-16 have product codes
            for col_letter in ['B', 'C']:
                total_lookup_cells += 1
                cell_val = ws[f'{col_letter}{row}'].value
                if cell_val is not None and isinstance(cell_val, str):
                    val_upper = cell_val.upper().replace(" ", "")
                    if 'VLOOKUP' in val_upper and 'CATALOG' in val_upper:
                        vlookup_count += 1

        if vlookup_count >= 12:
            # All 14 cells (7 rows x 2 cols) have VLOOKUP, allow minor misses
            print(f"PASS: Component 1 — {vlookup_count}/{total_lookup_cells} VLOOKUP formulas found (0.25 pts)")
            total_score += 0.25
        elif vlookup_count >= 7:
            partial = 0.25 * (vlookup_count / 14)
            print(f"PARTIAL: Component 1 — {vlookup_count}/{total_lookup_cells} VLOOKUP formulas ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {vlookup_count}/{total_lookup_cells} VLOOKUP formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ---------------------------------------------------------------
    # Component 2: Line total formulas in E column (0.20 points)
    # Initial: E10:E16 are None (empty)
    # Golden: formulas like =IF(A10="","",C10*D10) or =C10*D10
    # ---------------------------------------------------------------
    try:
        line_total_count = 0
        for row in range(10, 17):
            cell_val = ws[f'E{row}'].value
            if cell_val is not None and isinstance(cell_val, str):
                val_upper = cell_val.upper().replace(" ", "")
                # Check for multiplication pattern involving C and D columns
                if ('*' in val_upper) and (f'C{row}' in val_upper or f'D{row}' in val_upper):
                    line_total_count += 1

        if line_total_count >= 6:
            print(f"PASS: Component 2 — {line_total_count}/7 line total formulas found (0.20 pts)")
            total_score += 0.20
        elif line_total_count >= 3:
            partial = 0.20 * (line_total_count / 7)
            print(f"PARTIAL: Component 2 — {line_total_count}/7 line total formulas ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {line_total_count}/7 line total formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---------------------------------------------------------------
    # Component 3: Summary formulas - Subtotal, Discount, Shipping, Grand Total (0.25 points)
    # Initial: E22-E25 are all None
    # Golden: E22=SUM, E23=IF(discount), E24=SUMPRODUCT(shipping), E25=grand total
    # ---------------------------------------------------------------
    try:
        summary_score = 0.0

        # E22: Subtotal - should contain SUM formula referencing E column line totals
        e22 = ws['E22'].value
        if e22 is not None and isinstance(e22, str) and 'SUM' in e22.upper():
            print(f"  PASS: E22 Subtotal formula: {e22}")
            summary_score += 0.0625
        else:
            print(f"  FAIL: E22 Subtotal — expected SUM formula, found: {e22}")

        # E23: Discount - should contain IF formula with discount thresholds
        e23 = ws['E23'].value
        if e23 is not None and isinstance(e23, str):
            e23_upper = e23.upper().replace(" ", "")
            if 'IF' in e23_upper and ('0.1' in e23 or '0.05' in e23 or '10%' in e23 or '5%' in e23 or '*0.1' in e23_upper or '*0.05' in e23_upper):
                print(f"  PASS: E23 Discount formula: {e23}")
                summary_score += 0.0625
            else:
                print(f"  FAIL: E23 Discount — IF formula found but missing discount rates: {e23}")
        else:
            print(f"  FAIL: E23 Discount — expected IF formula, found: {e23}")

        # E24: Shipping - should reference weight and multiply by rate
        e24 = ws['E24'].value
        if e24 is not None and isinstance(e24, str):
            e24_upper = e24.upper().replace(" ", "")
            # Shipping formula should reference Catalog for weight and multiply by 0.5
            if ('0.5' in e24 or '0.50' in e24) and ('CATALOG' in e24_upper or 'WEIGHT' in e24_upper.replace(" ", "") or 'VLOOKUP' in e24_upper or 'SUMPRODUCT' in e24_upper):
                print(f"  PASS: E24 Shipping formula: {e24}")
                summary_score += 0.0625
            elif '0.5' in e24:
                # Has rate but maybe different structure
                print(f"  PARTIAL: E24 Shipping formula has $0.50 rate but different structure: {e24}")
                summary_score += 0.03
            else:
                print(f"  FAIL: E24 Shipping — expected weight*$0.50 formula, found: {e24}")
        else:
            print(f"  FAIL: E24 Shipping — expected formula, found: {e24}")

        # E25: Grand Total = Subtotal - Discount + Shipping
        e25 = ws['E25'].value
        if e25 is not None and isinstance(e25, str):
            e25_upper = e25.upper().replace(" ", "")
            # Should reference E22, E23, E24
            if 'E22' in e25_upper and 'E23' in e25_upper and 'E24' in e25_upper:
                print(f"  PASS: E25 Grand Total formula: {e25}")
                summary_score += 0.0625
            elif 'E22' in e25_upper:
                print(f"  PARTIAL: E25 Grand Total references E22 but missing others: {e25}")
                summary_score += 0.03
            else:
                print(f"  FAIL: E25 Grand Total — expected formula referencing E22-E24, found: {e25}")
        else:
            print(f"  FAIL: E25 Grand Total — expected formula, found: {e25}")

        print(f"  Component 3 subtotal: {summary_score:.4f}/0.25")
        total_score += summary_score
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ---------------------------------------------------------------
    # Component 4: Catalog sheet hidden (0.15 points)
    # Initial: visible, Golden: hidden
    # ---------------------------------------------------------------
    try:
        catalog_state = wb['Catalog'].sheet_state
        if catalog_state in ('hidden', 'veryHidden'):
            print(f"PASS: Component 4 — Catalog sheet is {catalog_state} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — Catalog sheet state is '{catalog_state}', expected 'hidden'")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ---------------------------------------------------------------
    # Component 5: Print area set on Order Form (0.15 points)
    # Initial: no print area, Golden: 'Order Form'!$A$1:$E$25
    # ---------------------------------------------------------------
    try:
        print_area = ws.print_area
        if print_area:
            # Accept any reasonable print area that covers the order form
            print_area_str = str(print_area)
            print(f"PASS: Component 5 — Print area set: {print_area_str} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 5 — No print area set")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score:.4f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
