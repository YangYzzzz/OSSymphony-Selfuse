"""
Reward Script: Unprotect the 'Report' sheet (no password)
Task ID: calc_ps_012
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): Sheet protection flag is disabled (protection.sheet == False)
  Component 2 (0.3): Password is cleared (protection.password is None)
  Component 3 (0.2): Data integrity — sheet still has data in A1:G30 AND protection is off
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_012'


def persist_app_state(domain: str):
    """Save any unsaved LibreOffice edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify that the 'Report' sheet has been unprotected.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that 'Report' sheet exists (precondition gate)
    if 'Report' not in wb.sheetnames:
        print(f"CRITICAL: 'Report' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Report']

    # Component 1: Sheet protection flag is disabled (0.5 points)
    # Initial state: protection.sheet == True
    # Golden state: protection.sheet == False
    try:
        sheet_protected = ws.protection.sheet
        if sheet_protected is False or sheet_protected is None:
            print(f"PASS: Component 1 — protection.sheet is {sheet_protected} (unprotected) (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — protection.sheet is {sheet_protected}, expected False")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Password is cleared (0.3 points)
    # Initial state: protection.password == 'CE4B' (hash)
    # Golden state: protection.password is None
    # Only award points if protection is also disabled (anchored to task change)
    try:
        pwd = ws.protection.password
        is_unprotected = (ws.protection.sheet is False or ws.protection.sheet is None)
        if is_unprotected and (pwd is None or pwd == ''):
            print(f"PASS: Component 2 — password is {repr(pwd)} and sheet is unprotected (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — password={repr(pwd)}, protection.sheet={ws.protection.sheet}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Data integrity with unprotection (0.2 points)
    # Verify data is still present AND sheet is unprotected
    # This compound check ensures we only score when the task change happened
    try:
        is_unprotected = (ws.protection.sheet is False or ws.protection.sheet is None)
        # Check header row
        headers = [ws.cell(row=1, column=c).value for c in range(1, 8)]
        expected_headers = ['Employee', 'Department', 'Region', 'Q1 Sales', 'Q2 Sales', 'Q3 Sales', 'Q4 Sales']
        headers_ok = headers == expected_headers
        # Check data extent
        has_data = ws.max_row >= 30 and ws.max_column >= 7
        # Check a sample cell to ensure data wasn't wiped
        sample_val = ws.cell(row=2, column=1).value
        sample_ok = sample_val is not None and str(sample_val).strip() != ''

        if is_unprotected and headers_ok and has_data and sample_ok:
            print(f"PASS: Component 3 — data intact (headers match, {ws.max_row} rows, {ws.max_column} cols) and sheet unprotected (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — unprotected={is_unprotected}, headers_ok={headers_ok}, has_data={has_data}, sample_ok={sample_ok}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist app state before verification
persist_app_state("libreoffice_calc")

# Run verification
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
