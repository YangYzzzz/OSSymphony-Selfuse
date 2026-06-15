"""
Initial Setup: Create a protected 'Report' sheet with empty password
Task ID: calc_ps_012
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.worksheet.protection import SheetProtection

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_012'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Report ---
    ws = wb.active
    ws.title = 'Report'

    # Headers
    headers = ['Employee', 'Department', 'Region', 'Q1 Sales', 'Q2 Sales', 'Q3 Sales', 'Q4 Sales']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows (A2:G30 = 29 data rows)
    data = [
        ['Sarah Chen', 'Engineering', 'West', 45230, 51200, 48700, 52100],
        ['Marcus Johnson', 'Marketing', 'East', 38500, 42100, 39800, 44600],
        ['Priya Patel', 'Sales', 'South', 67800, 72400, 69500, 75200],
        ['James O\'Brien', 'Finance', 'North', 29400, 31200, 30800, 33100],
        ['Aiko Tanaka', 'Engineering', 'West', 52100, 55800, 53400, 57200],
        ['Carlos Rivera', 'Marketing', 'South', 41300, 43700, 42500, 46800],
        ['Emma Larsson', 'Sales', 'East', 58900, 63200, 60100, 65400],
        ['David Kim', 'Finance', 'North', 34700, 36900, 35800, 38200],
        ['Fatima Al-Hassan', 'Engineering', 'South', 49800, 53100, 51200, 54700],
        ['Robert Williams', 'Marketing', 'West', 36200, 38800, 37500, 40100],
        ['Maria Santos', 'Sales', 'North', 62400, 66800, 64100, 69500],
        ['Alex Petrov', 'Finance', 'East', 31800, 33600, 32700, 35400],
        ['Wei Zhang', 'Engineering', 'South', 47500, 50800, 49200, 52600],
        ['Rachel Thompson', 'Marketing', 'North', 39100, 41500, 40300, 43200],
        ['Oluwaseun Adeyemi', 'Sales', 'West', 71200, 76500, 73800, 79100],
        ['Michael Brown', 'Finance', 'South', 28900, 30700, 29800, 32400],
        ['Sofia Kowalski', 'Engineering', 'East', 51400, 54900, 53100, 56800],
        ['Nathan Brooks', 'Marketing', 'North', 37800, 40200, 39000, 41700],
        ['Yuki Yamamoto', 'Sales', 'West', 64700, 69300, 66800, 72100],
        ['Isabelle Moreau', 'Finance', 'East', 33200, 35100, 34200, 36800],
        ['Ahmed Hassan', 'Engineering', 'North', 48300, 51600, 50100, 53400],
        ['Laura Garcia', 'Marketing', 'South', 40500, 43100, 41800, 44900],
        ['Benjamin Lee', 'Sales', 'West', 59800, 64100, 61900, 67200],
        ['Nina Volkov', 'Finance', 'North', 30600, 32400, 31500, 34100],
        ['Thomas Mueller', 'Engineering', 'East', 53700, 57200, 55400, 59100],
        ['Grace Okafor', 'Marketing', 'South', 42800, 45600, 44100, 47300],
        ['Christopher Park', 'Sales', 'North', 66100, 70800, 68400, 73700],
        ['Hannah Schmidt', 'Finance', 'West', 35400, 37600, 36500, 39200],
        ['Daniel Nguyen', 'Engineering', 'South', 50200, 53700, 52000, 55500],
    ]

    money_fmt = '$#,##0'
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c >= 4:  # monetary columns
                cell.number_format = money_fmt
                cell.alignment = Alignment(horizontal='right')

    # Set column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12

    # Protect the sheet with empty password
    ws.protection = SheetProtection(
        sheet=True,
        password='',
        formatCells=False,
        formatColumns=False,
        formatRows=False,
        insertColumns=False,
        insertRows=False,
        insertHyperlinks=False,
        deleteColumns=False,
        deleteRows=False,
        selectLockedCells=False,
        sort=False,
        autoFilter=False,
        pivotTables=False,
        selectUnlockedCells=False,
    )
    ws.protection.enable()

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
