"""
Reward Script: Create monthly sales totals by summing each rep's monthly figures.
Task ID: calc_sales_010
Domain: libreoffice_calc
Scoring:
  - Component 1: E2 contains SUM formula for Alice's Q1 total (0.25)
  - Component 2: E3 contains SUM formula for Bob's Q1 total (0.25)
  - Component 3: E4 contains SUM formula for Carol's Q1 total (0.25)
  - Component 4: E5 contains SUM formula for Dan's Q1 total (0.25)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_010'

# Ground truth: expected SUM formulas and their computed values
EXPECTED = {
    'E2': {'formula': '=SUM(B2:D2)', 'value': 131000},
    'E3': {'formula': '=SUM(B3:D3)', 'value': 162000},
    'E4': {'formula': '=SUM(B4:D4)', 'value': 113000},
    'E5': {'formula': '=SUM(B5:D5)', 'value': 145000},
}

COMPONENT_WEIGHT = 0.25


def check_sum_formula(ws, coord, expected_formula, expected_value):
    """
    Check if a cell contains a valid SUM formula for the correct range.
    Accepts the exact expected formula or any formula that would produce
    the correct result by summing the correct source cells.
    Returns True if the formula is correct.
    """
    val = ws[coord].value
    if val is None:
        return False, f"cell is empty (None)"

    if not isinstance(val, str) or not val.startswith('='):
        # Could be a hardcoded number — check if it matches the expected value
        # But we require a FORMULA, not a hardcoded value
        return False, f"not a formula, found: {val!r}"

    # Normalize for comparison: uppercase, strip spaces
    normalized = val.upper().replace(" ", "")
    expected_norm = expected_formula.upper().replace(" ", "")

    if normalized == expected_norm:
        return True, f"exact formula match: {val}"

    # Accept equivalent formulas that reference the same cells
    # e.g., =B2+C2+D2 instead of =SUM(B2:D2)
    # We verify by checking if the formula references the correct row
    row_num = coord[1:]  # e.g., '2' from 'E2'
    # Check if it references B, C, D of the same row
    has_b = f'B{row_num}' in normalized
    has_c = f'C{row_num}' in normalized
    has_d = f'D{row_num}' in normalized

    if has_b and has_c and has_d:
        return True, f"equivalent formula referencing correct cells: {val}"

    return False, f"formula does not reference expected cells: {val}"


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Monthly' sheet must exist
    if 'Monthly' not in wb.sheetnames:
        print(f"CRITICAL: 'Monthly' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Monthly']

    # Component 1: E2 contains SUM formula for Alice's Q1 total (0.25 points)
    try:
        passed, details = check_sum_formula(ws, 'E2', EXPECTED['E2']['formula'], EXPECTED['E2']['value'])
        if passed:
            print(f"PASS: Component 1 - E2 SUM formula for Alice ({details}) (0.25 pts)")
            total_score += COMPONENT_WEIGHT
        else:
            print(f"FAIL: Component 1 - E2 SUM formula for Alice: {details}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: E3 contains SUM formula for Bob's Q1 total (0.25 points)
    try:
        passed, details = check_sum_formula(ws, 'E3', EXPECTED['E3']['formula'], EXPECTED['E3']['value'])
        if passed:
            print(f"PASS: Component 2 - E3 SUM formula for Bob ({details}) (0.25 pts)")
            total_score += COMPONENT_WEIGHT
        else:
            print(f"FAIL: Component 2 - E3 SUM formula for Bob: {details}")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: E4 contains SUM formula for Carol's Q1 total (0.25 points)
    try:
        passed, details = check_sum_formula(ws, 'E4', EXPECTED['E4']['formula'], EXPECTED['E4']['value'])
        if passed:
            print(f"PASS: Component 3 - E4 SUM formula for Carol ({details}) (0.25 pts)")
            total_score += COMPONENT_WEIGHT
        else:
            print(f"FAIL: Component 3 - E4 SUM formula for Carol: {details}")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: E5 contains SUM formula for Dan's Q1 total (0.25 points)
    try:
        passed, details = check_sum_formula(ws, 'E5', EXPECTED['E5']['formula'], EXPECTED['E5']['value'])
        if passed:
            print(f"PASS: Component 4 - E5 SUM formula for Dan ({details}) (0.25 pts)")
            total_score += COMPONENT_WEIGHT
        else:
            print(f"FAIL: Component 4 - E5 SUM formula for Dan: {details}")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
