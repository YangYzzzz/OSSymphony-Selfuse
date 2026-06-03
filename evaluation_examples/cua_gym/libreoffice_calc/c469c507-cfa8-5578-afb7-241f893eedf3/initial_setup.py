"""
Initial Setup: Supply Chain Multi-Site Inventory Consolidation
Task ID: calc_ops_supply_chain_multi_site_inventory_074
Domain: libreoffice_calc

Creates an .xlsx file with 3 site sheets (Site-North, Site-Central, Site-South)
each containing 50 SKU records, plus a NetworkView sheet with the SKU list
pre-populated but all consolidation formulas empty (to be filled by the agent).
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_supply_chain_multi_site_inventory_074'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

# 50 realistic SKUs across product categories
SKUS = [
    ("SKU-0001", "Industrial Solvent A"),
    ("SKU-0002", "Steel Pipe 2in x 10ft"),
    ("SKU-0003", "Circuit Breaker 20A"),
    ("SKU-0004", "Safety Gloves L"),
    ("SKU-0005", "Hydraulic Fluid 5L"),
    ("SKU-0006", "Bearing Assembly 6205"),
    ("SKU-0007", "Filter Cartridge HEPA"),
    ("SKU-0008", "LED Panel 60W"),
    ("SKU-0009", "Cable Conduit 1in"),
    ("SKU-0010", "Pneumatic Valve 3/4in"),
    ("SKU-0011", "Lubricant Grease 1kg"),
    ("SKU-0012", "Stainless Bolt M10x50"),
    ("SKU-0013", "Pressure Gauge 100psi"),
    ("SKU-0014", "Motor Brushes Set"),
    ("SKU-0015", "Paint Primer 4L White"),
    ("SKU-0016", "Forklift Battery 48V"),
    ("SKU-0017", "Welding Rod E6013"),
    ("SKU-0018", "Safety Helmet Class E"),
    ("SKU-0019", "Stretch Wrap 20in"),
    ("SKU-0020", "Pallet Jack 5500lb"),
    ("SKU-0021", "Air Compressor Belt"),
    ("SKU-0022", "Drill Bit Set HSS"),
    ("SKU-0023", "Conveyor Belt Segment"),
    ("SKU-0024", "Electric Motor 1HP"),
    ("SKU-0025", "Coolant MixReady 20L"),
    ("SKU-0026", "Steel Shelf Bracket"),
    ("SKU-0027", "Hand Truck 600lb"),
    ("SKU-0028", "PP Strapping Roll"),
    ("SKU-0029", "Safety Goggles Anti-Fog"),
    ("SKU-0030", "Industrial Tape 2in"),
    ("SKU-0031", "Centrifugal Pump 1in"),
    ("SKU-0032", "Fuse Block 10-Circuit"),
    ("SKU-0033", "Chain Hoist 2T Manual"),
    ("SKU-0034", "Hex Key Set Metric"),
    ("SKU-0035", "Fire Extinguisher 5lb"),
    ("SKU-0036", "Toggle Clamp 200lb"),
    ("SKU-0037", "Conveyor Roller 3in"),
    ("SKU-0038", "Label Printer Ribbon"),
    ("SKU-0039", "PVC Elbow 90deg 2in"),
    ("SKU-0040", "Industrial Vacuum Bag"),
    ("SKU-0041", "Tool Box 24in Steel"),
    ("SKU-0042", "Cable Ties 12in Blk"),
    ("SKU-0043", "Pneumatic Cylinder 3in"),
    ("SKU-0044", "Anti-Static Floor Mat"),
    ("SKU-0045", "Oil Drain Pan 8qt"),
    ("SKU-0046", "Bench Vise 4in"),
    ("SKU-0047", "Spray Paint RAL7035"),
    ("SKU-0048", "Insulation Tape Elec"),
    ("SKU-0049", "Forklift Fork Extension"),
    ("SKU-0050", "Work Light LED 50W"),
]

# Min stock levels (uniform per SKU across all sites)
MIN_STOCKS = [
    20, 15, 30, 50, 25, 10, 40, 15, 35, 12,
    60, 100, 8,  25, 20, 5,  80, 40, 45, 2,
    30, 15, 6,  4,  18, 55, 3,  70, 60, 120,
    4,  10, 2,  20, 8,  25, 40, 90, 55, 200,
    3,  300, 4, 5,  15, 3,  30, 400, 2,  6,
]

# Unit costs per SKU (USD)
UNIT_COSTS = [
    12.50, 34.75, 18.90, 8.25,  22.00, 45.60, 28.30, 85.00, 9.40,  67.80,
    14.20, 0.85,  32.00, 16.50, 19.75, 485.00, 1.20, 24.00, 11.30, 2200.00,
    18.60, 45.90, 320.00, 780.00, 28.50, 4.30, 125.00, 22.80, 7.60, 2.40,
    890.00, 28.50, 450.00, 32.00, 58.00, 18.70, 14.80, 12.60, 5.90, 0.45,
    78.00, 0.08, 165.00, 42.00, 14.50, 94.00, 8.30, 1.80, 385.00, 68.00,
]

# Site-North quantities — mix of overstocked, normal, understocked
NORTH_QTYS = [
    85,  12,  95,  180, 42,  8,   110, 28,  90,  5,
    200, 350, 15,  80,  65,  12,  240, 145, 130, 4,
    95,  38,  8,   10,  55,  180, 7,   210, 195, 380,
    12,  32,  5,   65,  22,  80,  120, 280, 165, 620,
    8,   950, 10,  12,  48,  7,   95,  1250, 5,  18,
]

# Site-Central quantities
CENTRAL_QTYS = [
    45,  42,  28,  60,  85,  32,  25,  8,   14,  28,
    35,  85,  4,   12,  22,  3,   45,  20,  18,  1,
    12,  8,   3,   2,   8,   42,  1,   35,  18,  65,
    2,   8,   1,   8,   4,   10,  18,  42,  22,  185,
    2,   380, 2,   3,   8,   1,   12,  480, 1,   5,
]

# Site-South quantities
SOUTH_QTYS = [
    32,  65,  45,  95,  18,  25,  58,  42,  120, 18,
    48,  120, 22,  65,  45,  8,   95,  65,  82,  3,
    48,  28,  12,  6,   32,  95,  4,   145, 88,  210,
    5,   18,  3,   28,  14,  32,  55,  115, 78,  420,
    3,   580, 6,   8,   25,  4,   48,  820, 2,   12,
]


def create_initial():
    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ #
    # Sheet 1: Site-North
    # ------------------------------------------------------------------ #
    ws_north = wb.active
    ws_north.title = 'Site-North'
    headers = ['SKU', 'Product', 'Qty', 'Min Stock', 'Unit Cost']
    for col, h in enumerate(headers, 1):
        ws_north.cell(row=1, column=col, value=h)
    for i, (sku, product) in enumerate(SKUS):
        r = i + 2
        ws_north.cell(row=r, column=1, value=sku)
        ws_north.cell(row=r, column=2, value=product)
        ws_north.cell(row=r, column=3, value=NORTH_QTYS[i])
        ws_north.cell(row=r, column=4, value=MIN_STOCKS[i])
        ws_north.cell(row=r, column=5, value=UNIT_COSTS[i])

    # ------------------------------------------------------------------ #
    # Sheet 2: Site-Central
    # ------------------------------------------------------------------ #
    ws_central = wb.create_sheet('Site-Central')
    for col, h in enumerate(headers, 1):
        ws_central.cell(row=1, column=col, value=h)
    for i, (sku, product) in enumerate(SKUS):
        r = i + 2
        ws_central.cell(row=r, column=1, value=sku)
        ws_central.cell(row=r, column=2, value=product)
        ws_central.cell(row=r, column=3, value=CENTRAL_QTYS[i])
        ws_central.cell(row=r, column=4, value=MIN_STOCKS[i])
        ws_central.cell(row=r, column=5, value=UNIT_COSTS[i])

    # ------------------------------------------------------------------ #
    # Sheet 3: Site-South
    # ------------------------------------------------------------------ #
    ws_south = wb.create_sheet('Site-South')
    for col, h in enumerate(headers, 1):
        ws_south.cell(row=1, column=col, value=h)
    for i, (sku, product) in enumerate(SKUS):
        r = i + 2
        ws_south.cell(row=r, column=1, value=sku)
        ws_south.cell(row=r, column=2, value=product)
        ws_south.cell(row=r, column=3, value=SOUTH_QTYS[i])
        ws_south.cell(row=r, column=4, value=MIN_STOCKS[i])
        ws_south.cell(row=r, column=5, value=UNIT_COSTS[i])

    # ------------------------------------------------------------------ #
    # Sheet 4: NetworkView  (pre-filled SKU & Product in A & B; C:L empty)
    # ------------------------------------------------------------------ #
    ws_net = wb.create_sheet('NetworkView')
    net_headers = [
        'SKU',           # A
        'Product',       # B
        'North Qty',     # C  — empty (agent must fill cross-sheet refs)
        'Central Qty',   # D  — empty
        'South Qty',     # E  — empty
        'Total Network Qty',  # F — empty
        'Min Stock',     # G  — empty
        'Total Min Stock',    # H — empty
        'Network Balance',    # I — empty
        'Rebalance Flag',     # J — empty
        'Unit Cost',     # K  — empty
        'Network Value', # L  — empty
    ]
    for col, h in enumerate(net_headers, 1):
        ws_net.cell(row=1, column=col, value=h)

    # Pre-populate SKU and Product in columns A and B only
    for i, (sku, product) in enumerate(SKUS):
        r = i + 2
        ws_net.cell(row=r, column=1, value=sku)
        ws_net.cell(row=r, column=2, value=product)
        # Columns C through L intentionally left empty

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheets: Site-North, Site-Central, Site-South, NetworkView')
    print(f'  SKUs: 50 per site sheet')
    print(f'  NetworkView C:L columns: empty (to be filled by agent)')


create_initial()
