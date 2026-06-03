"""
Reward Script: Consolidate inventory data from three warehouse sites into a master NetworkView sheet
Task ID: calc_ops_supply_chain_multi_site_inventory_074
Domain: libreoffice_calc
Scoring:
  Component 1: Cross-site quantity references in C, D, E columns (0.25 pts)
  Component 2: Total network quantity formula in F column (0.15 pts)
  Component 3: Min stock pull (G) and total min stock formula H=G*3 (0.15 pts)
  Component 4: Network balance formula I=F-H (0.15 pts)
  Component 5: Rebalance flag formula in J column (REBALANCE logic) (0.15 pts)
  Component 6: Unit cost reference (K) and network value formula L=F*K (0.15 pts)
Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_supply_chain_multi_site_inventory_074'


def normalize_formula(f):
    """Normalize formula string for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: NetworkView sheet must exist
    if 'NetworkView' not in wb.sheetnames:
        print("FAIL: 'NetworkView' sheet not found in workbook")
        print(f"\nScore: 0.0/1.0")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['NetworkView']

    # Verify data rows span rows 2-51 (50 SKUs)
    # Columns A and B should already be populated (precondition, not scored)

    # Component 1: Cross-site quantity references in C, D, E columns (0.25 pts)
    # C2:C51 should reference 'Site-North'!C*, D2:D51 -> 'Site-Central'!C*, E2:E51 -> 'Site-South'!C*
    try:
        c_refs_ok = 0
        d_refs_ok = 0
        e_refs_ok = 0

        for row in range(2, 52):
            c_val = ws.cell(row=row, column=3).value  # Col C
            d_val = ws.cell(row=row, column=4).value  # Col D
            e_val = ws.cell(row=row, column=5).value  # Col E

            c_norm = normalize_formula(c_val)
            d_norm = normalize_formula(d_val)
            e_norm = normalize_formula(e_val)

            # C column: references Site-North qty column C
            if isinstance(c_val, str) and 'SITE-NORTH' in c_norm and f'!C{row}' in c_norm:
                c_refs_ok += 1

            # D column: references Site-Central qty column C
            if isinstance(d_val, str) and 'SITE-CENTRAL' in d_norm and f'!C{row}' in d_norm:
                d_refs_ok += 1

            # E column: references Site-South qty column C
            if isinstance(e_val, str) and 'SITE-SOUTH' in e_norm and f'!C{row}' in e_norm:
                e_refs_ok += 1

        # All 50 rows should have correct references
        if c_refs_ok == 50 and d_refs_ok == 50 and e_refs_ok == 50:
            print(f"PASS: Component 1 — All 50 rows have correct cross-site qty references "
                  f"(North: {c_refs_ok}, Central: {d_refs_ok}, South: {e_refs_ok}) (0.25 pts)")
            total_score += 0.25
        elif c_refs_ok > 0 or d_refs_ok > 0 or e_refs_ok > 0:
            print(f"FAIL: Component 1 — Partial cross-site refs: "
                  f"North={c_refs_ok}/50, Central={d_refs_ok}/50, South={e_refs_ok}/50")
        else:
            print(f"FAIL: Component 1 — No cross-site quantity references found in C/D/E columns")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Total network quantity formula in F column (0.15 pts)
    # F2:F51 should contain =C2+D2+E2 style formulas (sum of three site columns)
    try:
        f_formulas_ok = 0
        for row in range(2, 52):
            f_val = ws.cell(row=row, column=6).value  # Col F
            f_norm = normalize_formula(f_val)
            # Expected pattern: =C{row}+D{row}+E{row}
            expected = f'=C{row}+D{row}+E{row}'
            if isinstance(f_val, str) and normalize_formula(expected) in f_norm:
                f_formulas_ok += 1
            # Also accept variations like =D+C+E or similar
            elif isinstance(f_val, str) and re.search(
                    r'=.*C\d+.*[+].*D\d+.*[+].*E\d+|=.*[CD]\d+.*[+].*[CDE]\d+.*[+].*[CDE]\d+', f_norm):
                f_formulas_ok += 1

        if f_formulas_ok == 50:
            print(f"PASS: Component 2 — All 50 rows have total network qty formula (F=C+D+E) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 — Only {f_formulas_ok}/50 rows have correct F formula")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Min stock pull (G) and total min stock formula H=G*3 (0.15 pts)
    # G2:G51 references Min Stock from a site sheet (Site-North col D)
    # H2:H51 = G*3
    try:
        g_refs_ok = 0
        h_formulas_ok = 0

        for row in range(2, 52):
            g_val = ws.cell(row=row, column=7).value  # Col G
            h_val = ws.cell(row=row, column=8).value  # Col H

            g_norm = normalize_formula(g_val)
            h_norm = normalize_formula(h_val)

            # G column: references any site sheet's Min Stock (column D)
            if isinstance(g_val, str) and ('SITE-NORTH' in g_norm or 'SITE-CENTRAL' in g_norm or 'SITE-SOUTH' in g_norm):
                if f'!D{row}' in g_norm:
                    g_refs_ok += 1

            # H column: should be G*3
            expected_h = f'=G{row}*3'
            if isinstance(h_val, str) and normalize_formula(expected_h) in h_norm:
                h_formulas_ok += 1

        if g_refs_ok == 50 and h_formulas_ok == 50:
            print(f"PASS: Component 3 — All 50 rows have Min Stock reference (G) and Total Min Stock formula H=G*3 (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 — Min Stock refs: G={g_refs_ok}/50, H={h_formulas_ok}/50")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Network balance formula I=F-H (0.15 pts)
    # I2:I51 = F2-H2 (positive = overstocked, negative = understocked at network level)
    try:
        i_formulas_ok = 0
        for row in range(2, 52):
            i_val = ws.cell(row=row, column=9).value  # Col I
            i_norm = normalize_formula(i_val)
            expected_i = f'=F{row}-H{row}'
            if isinstance(i_val, str) and normalize_formula(expected_i) in i_norm:
                i_formulas_ok += 1

        if i_formulas_ok == 50:
            print(f"PASS: Component 4 — All 50 rows have Network Balance formula I=F-H (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — Only {i_formulas_ok}/50 rows have correct I formula")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Rebalance flag formula in J column (0.15 pts)
    # J2:J51 should contain IF formula that outputs 'REBALANCE' when any site is overstocked
    # while another is understocked (site qty > G*2 AND another site qty < G)
    try:
        j_formulas_ok = 0
        j_rebalance_present = 0
        j_g_reference_ok = 0

        for row in range(2, 52):
            j_val = ws.cell(row=row, column=10).value  # Col J
            j_norm = normalize_formula(j_val)

            if isinstance(j_val, str) and j_val.startswith('='):
                j_formulas_ok += 1
                # Check that it outputs "REBALANCE"
                if 'REBALANCE' in j_norm:
                    j_rebalance_present += 1
                # Check that it references G (min stock) for thresholds
                if f'G{row}' in j_norm:
                    j_g_reference_ok += 1

        if j_formulas_ok == 50 and j_rebalance_present == 50 and j_g_reference_ok == 50:
            print(f"PASS: Component 5 — All 50 rows have Rebalance Flag formula with REBALANCE text and G threshold (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 5 — J formulas: {j_formulas_ok}/50, REBALANCE text: {j_rebalance_present}/50, G ref: {j_g_reference_ok}/50")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Unit cost reference (K) and network value formula L=F*K (0.15 pts)
    # K2:K51 references Unit Cost from a site sheet (Site-North col E)
    # L2:L51 = F*K (total network inventory value)
    try:
        k_refs_ok = 0
        l_formulas_ok = 0

        for row in range(2, 52):
            k_val = ws.cell(row=row, column=11).value  # Col K
            l_val = ws.cell(row=row, column=12).value  # Col L

            k_norm = normalize_formula(k_val)
            l_norm = normalize_formula(l_val)

            # K column: references any site sheet's Unit Cost (column E)
            if isinstance(k_val, str) and ('SITE-NORTH' in k_norm or 'SITE-CENTRAL' in k_norm or 'SITE-SOUTH' in k_norm):
                if f'!E{row}' in k_norm:
                    k_refs_ok += 1

            # L column: should be F*K (network value)
            expected_l = f'=F{row}*K{row}'
            if isinstance(l_val, str) and normalize_formula(expected_l) in l_norm:
                l_formulas_ok += 1

        if k_refs_ok == 50 and l_formulas_ok == 50:
            print(f"PASS: Component 6 — All 50 rows have Unit Cost reference (K) and Network Value formula L=F*K (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 6 — Unit Cost refs: K={k_refs_ok}/50, Network Value: L={l_formulas_ok}/50")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
