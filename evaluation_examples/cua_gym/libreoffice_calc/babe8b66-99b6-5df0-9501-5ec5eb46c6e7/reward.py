"""
Reward Script: Create a pivot table from project hours data with BillableAmount calculated field
Task ID: calc_gcp_080
Domain: libreoffice_calc
Scoring:
  Component 1: PivotTable sheet exists (0.15)
  Component 2: Correct column headers (0.15)
  Component 3: All 6 projects present as rows (0.20)
  Component 4: BillableAmount values correct (0.25)
  Component 5: Sum of Hours values correct (0.15)
  Component 6: Grand Total row correct (0.10)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_080'

# Expected pivot data from task context and golden exploration
EXPECTED_PROJECTS = {
    'Project-A': {'hours': 320, 'billable': 38400},
    'Project-B': {'hours': 240.5, 'billable': 30612.5},
    'Project-C': {'hours': 274.5, 'billable': 34955},
    'Project-D': {'hours': 291, 'billable': 37812.5},
    'Project-E': {'hours': 317, 'billable': 42605},
    'Project-F': {'hours': 272.5, 'billable': 34052.5},
}

EXPECTED_GRAND_TOTAL_HOURS = 1715.5
EXPECTED_GRAND_TOTAL_BILLABLE = 218437.5


def persist_app_state(domain):
    """Save any unsaved LibreOffice state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: PivotTable sheet exists (0.15 points)
    # This FAILS on initial (only TimeEntries) and PASSES on golden
    try:
        pivot_sheet_name = next((sn for sn in wb.sheetnames if 'pivot' in sn.lower()), None)
        if pivot_sheet_name is not None:
            print(f"PASS: Component 1 -- PivotTable sheet found: '{pivot_sheet_name}' (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 -- No pivot table sheet found. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0  # No point continuing without pivot sheet
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[pivot_sheet_name]

    # Find header row and column mapping
    # The pivot table may start at different rows; scan for a row that has
    # "Project" AND at least one more recognized header in the same row
    header_row = None
    col_project = None
    col_hours = None
    col_billable = None

    for row_idx in range(1, min(ws.max_row + 1, 20)):
        row_vals = {}
        for col_idx in range(1, min(ws.max_column + 1, 20)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and isinstance(val, str):
                row_vals[col_idx] = val.strip()

        # Look for a row that has a "Project" column AND a hours/billable column
        tmp_project = None
        tmp_hours = None
        tmp_billable = None
        for col_idx, val in row_vals.items():
            val_lower = val.lower()
            # Match "Project" exactly (not "Project Hours Summary" or similar long titles)
            if val_lower == 'project':
                tmp_project = col_idx
            elif 'hour' in val_lower or 'sum' in val_lower:
                tmp_hours = col_idx
            if 'billable' in val_lower or 'amount' in val_lower:
                tmp_billable = col_idx

        if tmp_project and (tmp_hours or tmp_billable):
            header_row = row_idx
            col_project = tmp_project
            col_hours = tmp_hours
            col_billable = tmp_billable
            break

    # Component 2: Correct column headers (0.15 points)
    # Checks that the pivot table has the right structure with Project, Hours, and BillableAmount columns
    try:
        if header_row and col_project and col_hours and col_billable:
            print(f"PASS: Component 2 -- Headers found at row {header_row}: "
                  f"Project=col{col_project}, Hours=col{col_hours}, BillableAmount=col{col_billable} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 -- Missing headers. header_row={header_row}, "
                  f"col_project={col_project}, col_hours={col_hours}, col_billable={col_billable}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    if not (header_row and col_project and col_hours and col_billable):
        # Can't continue scoring without proper structure
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    # Read pivot data rows (after header, until empty or Grand Total)
    data_start = header_row + 1
    pivot_data = {}
    grand_total_row = None

    for row_idx in range(data_start, min(ws.max_row + 1, data_start + 20)):
        proj_val = ws.cell(row=row_idx, column=col_project).value
        if proj_val is None:
            continue
        proj_str = str(proj_val).strip()
        if 'grand' in proj_str.lower() or 'total' in proj_str.lower():
            grand_total_row = row_idx
            continue
        hours_val = ws.cell(row=row_idx, column=col_hours).value
        billable_val = ws.cell(row=row_idx, column=col_billable).value
        pivot_data[proj_str] = {
            'hours': hours_val,
            'billable': billable_val,
            'row': row_idx
        }

    # Component 3: All 6 projects present as rows (0.20 points)
    # Each project found earns proportional credit
    try:
        projects_found = 0
        for proj_name in EXPECTED_PROJECTS:
            if proj_name in pivot_data:
                projects_found += 1
            else:
                print(f"  MISS: Project '{proj_name}' not found in pivot data")

        if projects_found == 6:
            print(f"PASS: Component 3 -- All 6 projects found in pivot table (0.20 pts)")
            total_score += 0.20
        elif projects_found > 0:
            partial = round(0.20 * projects_found / 6, 4)
            print(f"PARTIAL: Component 3 -- {projects_found}/6 projects found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 -- No expected projects found. Found: {list(pivot_data.keys())}")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: BillableAmount values correct (0.25 points)
    # This is the key calculated field: Hours * HourlyRate aggregated per project
    try:
        billable_correct = 0
        tolerance = 1.0  # Allow small rounding differences
        for proj_name, expected in EXPECTED_PROJECTS.items():
            if proj_name in pivot_data:
                actual = pivot_data[proj_name]['billable']
                if actual is not None:
                    try:
                        if abs(float(actual) - expected['billable']) <= tolerance:
                            billable_correct += 1
                        else:
                            print(f"  MISMATCH: {proj_name} BillableAmount: expected={expected['billable']}, actual={actual}")
                    except (ValueError, TypeError):
                        print(f"  ERROR: {proj_name} BillableAmount not numeric: {actual}")

        if billable_correct == 6:
            print(f"PASS: Component 4 -- All 6 BillableAmount values correct (0.25 pts)")
            total_score += 0.25
        elif billable_correct > 0:
            partial = round(0.25 * billable_correct / 6, 4)
            print(f"PARTIAL: Component 4 -- {billable_correct}/6 BillableAmount values correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 -- No BillableAmount values match expected")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # Component 5: Sum of Hours values correct (0.15 points)
    try:
        hours_correct = 0
        tolerance = 0.5
        for proj_name, expected in EXPECTED_PROJECTS.items():
            if proj_name in pivot_data:
                actual = pivot_data[proj_name]['hours']
                if actual is not None:
                    try:
                        if abs(float(actual) - expected['hours']) <= tolerance:
                            hours_correct += 1
                        else:
                            print(f"  MISMATCH: {proj_name} Hours: expected={expected['hours']}, actual={actual}")
                    except (ValueError, TypeError):
                        print(f"  ERROR: {proj_name} Hours not numeric: {actual}")

        if hours_correct == 6:
            print(f"PASS: Component 5 -- All 6 Sum of Hours values correct (0.15 pts)")
            total_score += 0.15
        elif hours_correct > 0:
            partial = round(0.15 * hours_correct / 6, 4)
            print(f"PARTIAL: Component 5 -- {hours_correct}/6 Hours values correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 -- No Hours values match expected")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    # Component 6: Grand Total row present and correct (0.10 points)
    try:
        if grand_total_row is not None:
            gt_hours = ws.cell(row=grand_total_row, column=col_hours).value
            gt_billable = ws.cell(row=grand_total_row, column=col_billable).value
            gt_hours_ok = (gt_hours is not None and abs(float(gt_hours) - EXPECTED_GRAND_TOTAL_HOURS) <= 1.0)
            gt_billable_ok = (gt_billable is not None and abs(float(gt_billable) - EXPECTED_GRAND_TOTAL_BILLABLE) <= 1.0)

            if gt_hours_ok and gt_billable_ok:
                print(f"PASS: Component 6 -- Grand Total row correct (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 6 -- Grand Total row values incorrect "
                      f"(hours={gt_hours}, billable={gt_billable})")
        else:
            print(f"FAIL: Component 6 -- No Grand Total row found")
    except Exception as e:
        print(f"ERROR: Component 6 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
