"""
Initial Setup: Create a TimeEntries spreadsheet with 400 rows of project hours data.
Task ID: calc_gcp_080
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_080'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

# Seed for reproducibility
random.seed(42)

PROJECTS = ['Project-A', 'Project-B', 'Project-C', 'Project-D', 'Project-E', 'Project-F']

EMPLOYEES = [
    'Sarah Chen', 'Marcus Johnson', 'Priya Patel', 'James O\'Brien',
    'Lin Wei', 'Amara Okafor', 'Carlos Rivera', 'Emma Larsson',
    'David Kim', 'Fatima Al-Hassan', 'Thomas Mueller', 'Yuki Tanaka',
    'Rachel Green', 'Omar Benali', 'Sophie Martin', 'Raj Kapoor',
    'Anna Kowalski', 'Michael Torres', 'Leila Nazari', 'Hans Zimmer',
]

HOURLY_RATES = [75, 85, 95, 100, 110, 120, 125, 135, 150, 160, 175, 185, 200]

BILLABLE_OPTIONS = ['Yes', 'No']

# We need:
#   Project-A: total hours = 320, total BillableAmount (Hours * HourlyRate) = 38400
#   Average HourlyRate for Project-A = 38400 / 320 = 120
#
# Strategy: Assign ~67 entries to each project (400/6 ~ 67).
# For Project-A, carefully control hours and rates so they sum correctly.

def generate_date():
    """Generate a random date in 2025."""
    start = date(2025, 1, 1)
    delta = random.randint(0, 364)
    return start + timedelta(days=delta)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'TimeEntries'

    # Headers
    headers = ['EntryID', 'Date', 'Employee', 'Project', 'Hours', 'HourlyRate', 'Billable']
    # Wait - the context says: A=EntryID, B=Date, C=Project, D=Employee, E=Hours, F=HourlyRate, G=Billable
    headers = ['EntryID', 'Date', 'Project', 'Employee', 'Hours', 'HourlyRate', 'Billable']
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 10

    # Build entries per project
    # Project-A: 80 entries, hours must sum to 320, billable amount must sum to 38400
    # Other projects: distribute remaining 320 entries (~64 each)
    entries = []

    # --- Project-A entries (80 entries) ---
    # We need sum(hours) = 320 and sum(hours * rate) = 38400
    # Use a fixed rate of 120 for all Project-A entries: 320 * 120 = 38400
    pa_count = 80
    pa_hours_list = []
    remaining_hours = 320.0
    for i in range(pa_count - 1):
        h = random.choice([2.0, 3.0, 4.0, 5.0, 6.0])
        pa_hours_list.append(h)
    used = sum(pa_hours_list)
    # Adjust: we need exactly 320
    # Scale approach: just set specific values
    # Actually let me be more precise. Generate 79 values, then set the last one.
    # With 79 entries averaging 4, we get ~316, close enough to adjust.
    pa_hours_list = []
    for i in range(pa_count - 1):
        pa_hours_list.append(4.0)
    # Sum so far: 79 * 4 = 316, last entry = 4.0 => total = 320
    pa_hours_list.append(4.0)
    assert sum(pa_hours_list) == 320.0

    for h in pa_hours_list:
        entries.append({
            'project': 'Project-A',
            'hours': h,
            'rate': 120,  # All at $120 so 320 * 120 = 38400
            'employee': random.choice(EMPLOYEES),
            'date': generate_date(),
            'billable': random.choice(BILLABLE_OPTIONS),
        })

    # --- Other projects ---
    other_projects = ['Project-B', 'Project-C', 'Project-D', 'Project-E', 'Project-F']
    remaining_count = 400 - pa_count  # 320 entries
    per_project = remaining_count // len(other_projects)  # 64 each

    for idx, proj in enumerate(other_projects):
        count = per_project if idx < len(other_projects) - 1 else (remaining_count - per_project * (len(other_projects) - 1))
        for _ in range(count):
            h = random.choice([0.5, 1.0, 1.5, 2.0, 2.5, 3.0, 3.5, 4.0, 4.5, 5.0, 5.5, 6.0, 6.5, 7.0, 7.5, 8.0])
            entries.append({
                'project': proj,
                'hours': h,
                'rate': random.choice(HOURLY_RATES),
                'employee': random.choice(EMPLOYEES),
                'date': generate_date(),
                'billable': random.choice(BILLABLE_OPTIONS),
            })

    # Shuffle entries
    random.shuffle(entries)

    # Write data rows
    for r, entry in enumerate(entries, 2):
        ws.cell(row=r, column=1, value=r - 1)  # EntryID
        ws.cell(row=r, column=2, value=entry['date']).number_format = 'YYYY-MM-DD'
        ws.cell(row=r, column=3, value=entry['project'])
        ws.cell(row=r, column=4, value=entry['employee'])
        ws.cell(row=r, column=5, value=entry['hours'])
        ws.cell(row=r, column=6, value=entry['rate']).number_format = '$#,##0'
        ws.cell(row=r, column=7, value=entry['billable'])

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Total entries: {len(entries)}')

    # Verify Project-A totals
    pa_entries = [e for e in entries if e['project'] == 'Project-A']
    pa_hours = sum(e['hours'] for e in pa_entries)
    pa_billable = sum(e['hours'] * e['rate'] for e in pa_entries)
    print(f'Project-A: {len(pa_entries)} entries, hours={pa_hours}, billable_amount={pa_billable}')


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


create_initial()
launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')
