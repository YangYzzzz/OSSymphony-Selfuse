"""
Initial Setup: Two-way INDEX/MATCH/MATCH lookup in a sales matrix
Task ID: calc_lf_001
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_001'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: SalesMatrix ---
    ws = wb.active
    ws.title = 'SalesMatrix'

    # Column headers (quarters) in row 1
    # A1 is intentionally empty
    ws.cell(row=1, column=2, value='Q1')
    ws.cell(row=1, column=3, value='Q2')
    ws.cell(row=1, column=4, value='Q3')
    ws.cell(row=1, column=5, value='Q4')

    # Product rows
    products = [
        ['Product A', 12000, 14500, 13200, 15800],
        ['Product B', 9800, 11200, 10500, 12300],
        ['Product C', 18700, 20100, 19400, 22600],
    ]
    for r, row_data in enumerate(products, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Lookup area in columns G-I
    ws.cell(row=1, column=7, value='Product')   # G1
    ws.cell(row=2, column=7, value='Product B')  # G2
    ws.cell(row=1, column=8, value='Quarter')    # H1
    ws.cell(row=2, column=8, value='Q3')         # H2
    ws.cell(row=1, column=9, value='Result')     # I1
    # I2 is intentionally left empty - this is where the agent must enter the formula

    # Light formatting for readability
    header_font = Font(bold=True)
    for col in range(2, 6):
        ws.cell(row=1, column=col).font = header_font
    for row in range(2, 5):
        ws.cell(row=row, column=1).font = header_font
    for col in [7, 8, 9]:
        ws.cell(row=1, column=col).font = header_font

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 14
    for col_letter in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col_letter].width = 10
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
