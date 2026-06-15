"""
Reward Script: Two-way INDEX/MATCH/MATCH lookup for sales figure
Task ID: calc_lf_001
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): I2 contains a formula (not empty/literal)
  Component 2 (0.3): Formula uses INDEX with nested MATCH functions
  Component 3 (0.3): Formula references correct ranges for two-way lookup
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_001'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: SalesMatrix sheet must exist
    if 'SalesMatrix' not in wb.sheetnames:
        print(f"CRITICAL: 'SalesMatrix' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['SalesMatrix']
    i2_value = ws['I2'].value

    # Component 1: I2 contains a formula (0.4 points)
    # This FAILS on initial (I2 is None) and PASSES on golden (I2 has formula)
    try:
        if i2_value is not None and isinstance(i2_value, str) and i2_value.strip().startswith('='):
            print(f"PASS: Component 1 -- I2 contains a formula: {i2_value} (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 -- I2 does not contain a formula. Value: {repr(i2_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Formula uses INDEX with nested MATCH (0.3 points)
    # Verify the formula contains INDEX(...MATCH...MATCH...) pattern
    try:
        if i2_value is not None and isinstance(i2_value, str):
            formula_upper = i2_value.upper().replace(" ", "")
            has_index = "INDEX(" in formula_upper
            # Count MATCH occurrences -- need at least 2 for two-way lookup
            match_count = formula_upper.count("MATCH(")
            if has_index and match_count >= 2:
                print(f"PASS: Component 2 -- Formula uses INDEX with {match_count} MATCH functions (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 -- Expected INDEX with 2 MATCH calls. "
                      f"INDEX present: {has_index}, MATCH count: {match_count}")
        else:
            print(f"FAIL: Component 2 -- I2 is not a formula string. Value: {repr(i2_value)}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Formula references correct ranges for two-way lookup (0.3 points)
    # Expected: =INDEX(B2:E4, MATCH(G2,A2:A4,0), MATCH(H2,B1:E1,0))
    # Check: data range includes the sales matrix, row lookup references product names,
    # column lookup references quarter headers
    try:
        if i2_value is not None and isinstance(i2_value, str):
            formula_clean = i2_value.upper().replace(" ", "")

            # Check data array covers the sales data area (B2:E4 or equivalent)
            # The INDEX first arg should be a range covering rows 2-4 and cols B-E
            has_data_range = bool(re.search(r'INDEX\(B2:E4', formula_clean))

            # Check row lookup references product names column (A2:A4)
            has_row_lookup = bool(re.search(r'MATCH\(G2,A2:A4', formula_clean))

            # Check column lookup references quarter headers (B1:E1)
            has_col_lookup = bool(re.search(r'MATCH\(H2,B1:E1', formula_clean))

            checks_passed = sum([has_data_range, has_row_lookup, has_col_lookup])

            if checks_passed == 3:
                print(f"PASS: Component 3 -- Formula references correct ranges "
                      f"(data=B2:E4, rows=A2:A4, cols=B1:E1) (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 3 -- Range checks: data_range={has_data_range}, "
                      f"row_lookup={has_row_lookup}, col_lookup={has_col_lookup}")
                # Award partial credit: 0.1 per correct range reference
                partial = checks_passed * 0.1
                if partial > 0:
                    print(f"  Partial credit: {partial} pts for {checks_passed}/3 correct ranges")
                    total_score += partial
        else:
            print(f"FAIL: Component 3 -- I2 is not a formula string. Value: {repr(i2_value)}")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
