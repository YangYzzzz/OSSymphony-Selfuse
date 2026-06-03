"""
Initial Setup: Create a workbook with Annual Summary and Q1-Q4 sheets (no tab colors set)
Task ID: calc_sht_tabcolor_002
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_sht_tabcolor_002'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ #
    # Sheet 1: Annual Summary — P&L dashboard
    # ------------------------------------------------------------------ #
    ws_annual = wb.active
    ws_annual.title = 'Annual Summary'

    # Header row
    annual_headers = ['Category', 'Q1', 'Q2', 'Q3', 'Q4', 'Annual Total']
    for col, h in enumerate(annual_headers, 1):
        cell = ws_annual.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='center')

    # P&L rows
    annual_data = [
        ['Revenue',          1_250_400, 1_380_750, 1_520_300, 1_695_200, 5_846_650],
        ['Cost of Goods',     750_240,   828_450,   912_180, 1_017_120, 3_507_990],
        ['Gross Profit',      500_160,   552_300,   608_120,   678_080, 2_338_660],
        ['Operating Expenses',185_300,   198_750,   204_600,   221_900,   810_550],
        ['EBITDA',            314_860,   353_550,   403_520,   456_180, 1_528_110],
        ['Depreciation',       28_000,    28_000,    28_000,    28_000,   112_000],
        ['EBIT',              286_860,   325_550,   375_520,   428_180, 1_416_110],
        ['Interest Expense',   12_400,    12_100,    11_800,    11_500,    47_800],
        ['Pre-Tax Income',    274_460,   313_450,   363_720,   416_680, 1_368_310],
        ['Tax (25%)',          68_615,    78_363,    90_930,   104_170,   342_078],
        ['Net Income',        205_845,   235_088,   272_790,   312_510, 1_026_233],
    ]

    for r, row_data in enumerate(annual_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_annual.cell(row=r, column=c, value=val)

    ws_annual.column_dimensions['A'].width = 22
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws_annual.column_dimensions[col_letter].width = 15

    # ------------------------------------------------------------------ #
    # Helper: create one quarterly sheet with 80 rows of financial data
    # ------------------------------------------------------------------ #
    departments = [
        'Engineering', 'Marketing', 'Sales', 'Operations',
        'HR', 'Finance', 'Customer Success', 'Product',
        'Legal', 'IT'
    ]
    categories = [
        'Salaries', 'Software Licenses', 'Travel & Entertainment',
        'Office Supplies', 'Advertising', 'Professional Services',
        'Equipment', 'Training'
    ]

    def make_quarterly_sheet(wb, sheet_name, base_revenue, base_expense_factor):
        ws = wb.create_sheet(sheet_name)

        # Header
        headers = ['Line Item', 'Department', 'Category', 'Jan', 'Feb', 'Mar', 'Total', 'vs Budget']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='center')

        # Generate 80 rows of realistic data
        import random
        random.seed(hash(sheet_name) % 1000)

        row_num = 2
        dept_cycle = departments * 8   # 80 entries
        cat_cycle  = categories  * 10  # repeat to cover 80 rows

        for i in range(80):
            dept = dept_cycle[i]
            cat  = cat_cycle[i]
            line_item = f'{dept} – {cat}'

            jan = round(base_revenue * base_expense_factor * random.uniform(0.8, 1.3) / 80, 2)
            feb = round(jan * random.uniform(0.92, 1.08), 2)
            mar = round(jan * random.uniform(0.92, 1.10), 2)
            total = round(jan + feb + mar, 2)
            budget_pct = round(random.uniform(-0.12, 0.18), 4)

            ws.cell(row=row_num, column=1, value=line_item)
            ws.cell(row=row_num, column=2, value=dept)
            ws.cell(row=row_num, column=3, value=cat)
            ws.cell(row=row_num, column=4, value=jan)
            ws.cell(row=row_num, column=5, value=feb)
            ws.cell(row=row_num, column=6, value=mar)
            ws.cell(row=row_num, column=7, value=total)
            ws.cell(row=row_num, column=8, value=budget_pct)
            row_num += 1

        # Column widths
        ws.column_dimensions['A'].width = 38
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 24
        for col_letter in ['D', 'E', 'F', 'G']:
            ws.column_dimensions[col_letter].width = 13
        ws.column_dimensions['H'].width = 12

        return ws

    # ------------------------------------------------------------------ #
    # Create Q1 – Q4 sheets (NO tab colors)
    # ------------------------------------------------------------------ #
    make_quarterly_sheet(wb, 'Q1', 1_250_400, 0.60)
    make_quarterly_sheet(wb, 'Q2', 1_380_750, 0.60)
    make_quarterly_sheet(wb, 'Q3', 1_520_300, 0.60)
    make_quarterly_sheet(wb, 'Q4', 1_695_200, 0.60)

    # ------------------------------------------------------------------ #
    # Ensure NO tab colors are set on any sheet
    # ------------------------------------------------------------------ #
    for ws in wb.worksheets:
        ws.sheet_properties.tabColor = None

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheets: {wb.sheetnames}')
    for ws in wb.worksheets:
        tc = ws.sheet_properties.tabColor
        print(f'  {ws.title}: tabColor={tc}')


create_initial()
