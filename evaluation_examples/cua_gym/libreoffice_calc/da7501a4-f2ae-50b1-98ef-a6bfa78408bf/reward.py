"""
Reward Script: Color-code quarterly sheets with specified tab colors
Task ID: calc_sht_tabcolor_002
Domain: libreoffice_calc
Scoring:
  Component 1: Q1 tab color == #70AD47 (green)       — 0.25 pts
  Component 2: Q2 tab color == #FFC000 (yellow-orange) — 0.25 pts
  Component 3: Q3 tab color == #FF0000 (red)          — 0.25 pts
  Component 4: Q4 tab color == #4472C4 (blue)         — 0.25 pts
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_sht_tabcolor_002'

# Expected tab colors in 8-char ARGB format (as stored by openpyxl)
EXPECTED_COLORS = {
    'Q1': 'FF70AD47',   # green
    'Q2': 'FFFFC000',   # yellow-orange
    'Q3': 'FFFF0000',   # red
    'Q4': 'FF4472C4',   # blue
}

# Each quarterly sheet contributes equally to the score
POINTS_PER_SHEET = 0.25


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Checks that Q1-Q4 sheet tabs are set to the required colors.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition gate: file must load
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: required sheets must exist
    required_sheets = ['Q1', 'Q2', 'Q3', 'Q4', 'Annual Summary']
    for sheet_name in required_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"CRITICAL: Sheet '{sheet_name}' not found in workbook.")
            print(f"  Present sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0

    # Component 1: Q1 tab color == #70AD47 (green) (0.25 points)
    try:
        ws_q1 = wb['Q1']
        tab_color = ws_q1.sheet_properties.tabColor
        if tab_color is not None and tab_color.rgb is not None:
            actual_rgb = tab_color.rgb.upper()
            expected_rgb = EXPECTED_COLORS['Q1'].upper()
            if actual_rgb == expected_rgb:
                print(f"PASS: Component 1 — Q1 tab color = {actual_rgb} (expected {expected_rgb}) ({POINTS_PER_SHEET} pts)")
                total_score += POINTS_PER_SHEET
            else:
                print(f"FAIL: Component 1 — Q1 tab color expected {expected_rgb}, found {actual_rgb}")
        else:
            print(f"FAIL: Component 1 — Q1 tab color is None (no color set)")
    except Exception as e:
        print(f"ERROR: Component 1 — checking Q1 tab color: {e}")

    # Component 2: Q2 tab color == #FFC000 (yellow-orange) (0.25 points)
    try:
        ws_q2 = wb['Q2']
        tab_color = ws_q2.sheet_properties.tabColor
        if tab_color is not None and tab_color.rgb is not None:
            actual_rgb = tab_color.rgb.upper()
            expected_rgb = EXPECTED_COLORS['Q2'].upper()
            if actual_rgb == expected_rgb:
                print(f"PASS: Component 2 — Q2 tab color = {actual_rgb} (expected {expected_rgb}) ({POINTS_PER_SHEET} pts)")
                total_score += POINTS_PER_SHEET
            else:
                print(f"FAIL: Component 2 — Q2 tab color expected {expected_rgb}, found {actual_rgb}")
        else:
            print(f"FAIL: Component 2 — Q2 tab color is None (no color set)")
    except Exception as e:
        print(f"ERROR: Component 2 — checking Q2 tab color: {e}")

    # Component 3: Q3 tab color == #FF0000 (red) (0.25 points)
    try:
        ws_q3 = wb['Q3']
        tab_color = ws_q3.sheet_properties.tabColor
        if tab_color is not None and tab_color.rgb is not None:
            actual_rgb = tab_color.rgb.upper()
            expected_rgb = EXPECTED_COLORS['Q3'].upper()
            if actual_rgb == expected_rgb:
                print(f"PASS: Component 3 — Q3 tab color = {actual_rgb} (expected {expected_rgb}) ({POINTS_PER_SHEET} pts)")
                total_score += POINTS_PER_SHEET
            else:
                print(f"FAIL: Component 3 — Q3 tab color expected {expected_rgb}, found {actual_rgb}")
        else:
            print(f"FAIL: Component 3 — Q3 tab color is None (no color set)")
    except Exception as e:
        print(f"ERROR: Component 3 — checking Q3 tab color: {e}")

    # Component 4: Q4 tab color == #4472C4 (blue) (0.25 points)
    try:
        ws_q4 = wb['Q4']
        tab_color = ws_q4.sheet_properties.tabColor
        if tab_color is not None and tab_color.rgb is not None:
            actual_rgb = tab_color.rgb.upper()
            expected_rgb = EXPECTED_COLORS['Q4'].upper()
            if actual_rgb == expected_rgb:
                print(f"PASS: Component 4 — Q4 tab color = {actual_rgb} (expected {expected_rgb}) ({POINTS_PER_SHEET} pts)")
                total_score += POINTS_PER_SHEET
            else:
                print(f"FAIL: Component 4 — Q4 tab color expected {expected_rgb}, found {actual_rgb}")
        else:
            print(f"FAIL: Component 4 — Q4 tab color is None (no color set)")
    except Exception as e:
        print(f"ERROR: Component 4 — checking Q4 tab color: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
