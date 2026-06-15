"""
Reward Script: Sales commission lookup via VLOOKUP
Task ID: calc_sales_commission_lookup_006
Domain: libreoffice_calc
Scoring:
  - Component 1: VLOOKUP formulas in E2:E21 referencing Rates sheet (0.40 pts)
  - Component 2: Commission formulas (=D*E) in F2:F21 (0.30 pts)
  - Component 3: Summary row 22 — 'TOTAL' in A22 and SUM(F2:F21) in F22 (0.20 pts)
  - Component 4: AVERAGE formula in E22 (0.10 pts)
  NOTE: Rates sheet integrity is used as a precondition gate, not a scoring component,
        because the Rates sheet is already correct in the initial file.
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_sales_commission_lookup_006'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify RepCommissions sheet exists
    if 'RepCommissions' not in wb.sheetnames:
        print("CRITICAL: 'RepCommissions' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['RepCommissions']

    # Component 1: VLOOKUP formulas in E2:E21 referencing Rates sheet (0.40 points)
    # Each rep's base rate should be pulled from the Rates sheet via VLOOKUP
    # Initial state: E2:E21 are all None → any VLOOKUP formula indicates the task was done
    try:
        vlookup_count = 0
        for row in range(2, 22):
            e_val = ws.cell(row=row, column=5).value
            if (isinstance(e_val, str) and
                    'VLOOKUP' in e_val.upper() and
                    'RATES' in e_val.upper()):
                vlookup_count += 1

        if vlookup_count == 20:
            print(f"PASS: Component 1 — All 20 rows have VLOOKUP formula referencing Rates sheet (0.40 pts)")
            total_score += 0.40
        elif vlookup_count >= 10:
            partial = round(0.40 * vlookup_count / 20, 4)
            print(f"PARTIAL: Component 1 — {vlookup_count}/20 rows have VLOOKUP formula ({partial:.4f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {vlookup_count}/20 rows have VLOOKUP formula referencing Rates sheet")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Commission formulas in F2:F21 (0.30 points)
    # Each commission = Total Sales * Base Rate (=D*E formula)
    # Initial state: F2:F21 are all None → any D*E formula indicates the task was done
    try:
        commission_count = 0
        for row in range(2, 22):
            f_val = ws.cell(row=row, column=6).value
            if isinstance(f_val, str):
                f_upper = f_val.upper().replace(' ', '')
                # Check for pattern =D{row}*E{row} or =E{row}*D{row}
                if (f_upper == f'=D{row}*E{row}' or
                        f_upper == f'=E{row}*D{row}'):
                    commission_count += 1

        if commission_count == 20:
            print(f"PASS: Component 2 — All 20 rows have commission formula =D*E (0.30 pts)")
            total_score += 0.30
        elif commission_count >= 10:
            partial = round(0.30 * commission_count / 20, 4)
            print(f"PARTIAL: Component 2 — {commission_count}/20 rows have commission formula ({partial:.4f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {commission_count}/20 rows have commission formula =D*E")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Summary row — 'TOTAL' label in A22 and SUM formula in F22 (0.20 points)
    # Initial state: row 22 is completely empty → any TOTAL label + SUM formula means task was done
    try:
        a22 = ws.cell(row=22, column=1).value
        f22 = ws.cell(row=22, column=6).value

        has_total_label = (isinstance(a22, str) and a22.strip().upper() == 'TOTAL')
        has_sum_formula = (isinstance(f22, str) and
                           'SUM' in f22.upper() and
                           'F2' in f22.upper() and
                           'F21' in f22.upper())

        if has_total_label and has_sum_formula:
            print(f"PASS: Component 3 — A22='TOTAL' and F22 has SUM(F2:F21) formula (0.20 pts)")
            total_score += 0.20
        elif has_total_label and not has_sum_formula:
            print(f"PARTIAL: Component 3 — A22='TOTAL' but F22 SUM formula missing/incorrect. F22={repr(f22)} (0.08 pts)")
            total_score += 0.08
        elif not has_total_label and has_sum_formula:
            print(f"PARTIAL: Component 3 — F22 has SUM formula but A22 label missing/incorrect. A22={repr(a22)} (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 3 — A22={repr(a22)}, F22={repr(f22)} (expected 'TOTAL' and SUM formula)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: AVERAGE formula in E22 (0.10 points)
    # Summary row should include average rate across all reps
    # Initial state: E22 is None → any AVERAGE formula means task was done
    try:
        e22 = ws.cell(row=22, column=5).value

        has_avg_formula = (isinstance(e22, str) and
                           'AVERAGE' in e22.upper() and
                           'E2' in e22.upper() and
                           'E21' in e22.upper())

        if has_avg_formula:
            print(f"PASS: Component 4 — E22 has AVERAGE(E2:E21) formula (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 — E22={repr(e22)} (expected AVERAGE formula referencing E2:E21)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Precondition check: Rates sheet integrity (informational only, not scored)
    # The Rates sheet is already correct in the initial file, so checking it here
    # would award points for a pre-existing condition — not scored.
    try:
        if 'Rates' not in wb.sheetnames:
            print("WARNING: 'Rates' sheet is missing (task requires it for VLOOKUP)")
        else:
            ws_rates = wb['Rates']
            rates_row_count = sum(1 for row in range(2, 8)
                                  if ws_rates.cell(row=row, column=1).value is not None)
            print(f"INFO: Rates sheet present with {rates_row_count} territory rows (precondition check only, not scored)")
    except Exception as e:
        print(f"INFO: Could not verify Rates sheet — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
