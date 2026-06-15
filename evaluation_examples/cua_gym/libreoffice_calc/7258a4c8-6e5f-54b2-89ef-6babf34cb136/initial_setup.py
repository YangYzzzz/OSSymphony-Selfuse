"""
Initial Setup: Sales Commission VLOOKUP Task
Task ID: calc_sales_commission_lookup_006
Domain: libreoffice_calc

Creates a workbook with:
  - 'RepCommissions' sheet: 20 sales reps with Rep ID, Rep Name, Territory,
    Total Sales. Columns E (Base Rate) and F (Commission) are intentionally empty.
  - 'Rates' sheet: Territory-to-base-rate lookup table (5 territories sorted
    alphabetically).
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_commission_lookup_006'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ #
    # Sheet 1: RepCommissions
    # ------------------------------------------------------------------ #
    ws1 = wb.active
    ws1.title = 'RepCommissions'

    # Headers (Row 1)
    headers = ['Rep ID', 'Rep Name', 'Territory', 'Total Sales', 'Base Rate', 'Commission']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Column widths
    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 12
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 14

    # 20 reps data — realistic names, territories distributed
    # Territories: Central, East, North, South, West
    reps = [
        ('R001', 'Sarah Chen',         'North',   185420.50),
        ('R002', 'Marcus Johnson',      'South',   142890.75),
        ('R003', 'Priya Nair',          'East',    209340.00),
        ('R004', 'Derek Williams',      'West',    176500.25),
        ('R005', 'Amanda Torres',       'Central', 131780.90),
        ('R006', 'Kevin Park',          'North',   198650.40),
        ('R007', 'Lisa Monroe',         'South',   167230.60),
        ('R008', 'James Okafor',        'East',    224100.80),
        ('R009', 'Stephanie Reeves',    'West',    149760.15),
        ('R010', 'Carlos Mendez',       'Central', 118940.70),
        ('R011', 'Rachel Kim',          'North',   212560.35),
        ('R012', 'Thomas Brennan',      'South',   155380.55),
        ('R013', 'Aisha Patel',         'East',    189900.20),
        ('R014', 'Brandon Scott',       'West',    203410.65),
        ('R015', 'Megan Walsh',         'Central', 127650.45),
        ('R016', 'Liam Nguyen',         'North',   241780.00),
        ('R017', 'Olivia Hartman',      'South',   138270.85),
        ('R018', 'Samuel Diaz',         'East',    196450.30),
        ('R019', 'Natalie Foster',      'West',    162030.75),
        ('R020', 'Daniel Richardson',   'Central', 144890.10),
    ]

    for r, (rep_id, name, territory, sales) in enumerate(reps, 2):
        ws1.cell(row=r, column=1, value=rep_id)
        ws1.cell(row=r, column=2, value=name)
        ws1.cell(row=r, column=3, value=territory)
        cell_d = ws1.cell(row=r, column=4, value=sales)
        cell_d.number_format = '#,##0.00'
        # Columns E (Base Rate) and F (Commission) intentionally left empty

    # ------------------------------------------------------------------ #
    # Sheet 2: Rates (lookup table)
    # ------------------------------------------------------------------ #
    ws2 = wb.create_sheet('Rates')

    # Headers
    ws2.cell(row=1, column=1, value='Territory').font = Font(bold=True)
    ws2.cell(row=1, column=2, value='Base Rate').font = Font(bold=True)

    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 12

    # 5 territories sorted alphabetically with rates 0.06 – 0.11
    rates = [
        ('Central', 0.07),
        ('East',    0.09),
        ('North',   0.10),
        ('South',   0.06),
        ('West',    0.11),
    ]

    for r, (territory, rate) in enumerate(rates, 2):
        ws2.cell(row=r, column=1, value=territory)
        cell_rate = ws2.cell(row=r, column=2, value=rate)
        cell_rate.number_format = '0.00%'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
