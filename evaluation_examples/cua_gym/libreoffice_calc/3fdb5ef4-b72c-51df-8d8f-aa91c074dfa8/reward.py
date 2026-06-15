"""
Reward Script: Tabulate course evaluation survey results
Task ID: calc_edu_survey_tabulation_015
Domain: libreoffice_calc
Scoring:
  Component 1: AVERAGE formulas in B49:B53 for each question column (0.35 pts)
  Component 2: COUNTIF/45 formulas in C49:C53 for % satisfied (0.25 pts)
  Component 3: Number formats — B49:B53 as 0.00, C49:C53 as percentage (0.15 pts)
  Component 4: Radar chart present on Survey sheet (0.15 pts)
  Component 5: Chart title is 'Course Evaluation Radar' (0.10 pts)
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_survey_tabulation_015'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Survey' sheet must exist
    if 'Survey' not in wb.sheetnames:
        print("CRITICAL: 'Survey' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Survey']

    # Map of question rows to their source columns (B=2, C=3, D=4, E=5, F=6)
    # B49 -> AVERAGE(B2:B46), B50 -> AVERAGE(C2:C46), ...
    question_cols = {
        49: ('B', 2),  # row 49, source col B (col index 2)
        50: ('C', 3),  # row 50, source col C (col index 3)
        51: ('D', 4),  # row 51, source col D (col index 4)
        52: ('E', 5),  # row 52, source col E (col index 5)
        53: ('F', 6),  # row 53, source col F (col index 6)
    }

    # Component 1: AVERAGE formulas in B49:B53 (0.35 points)
    # Each question row must have an AVERAGE formula referencing the correct column (rows 2-46)
    # FAILS on initial (all None), PASSES on golden
    try:
        avg_ok_count = 0
        for row_num, (col_letter, col_idx) in question_cols.items():
            cell = ws.cell(row=row_num, column=2)  # column B
            val = cell.value
            if val is None:
                print(f"FAIL: Component 1 — B{row_num} is empty (expected AVERAGE formula)")
                continue
            if not isinstance(val, str):
                print(f"FAIL: Component 1 — B{row_num} has non-formula value: {repr(val)}")
                continue
            val_upper = val.upper().replace(' ', '')
            # Check it's an AVERAGE formula referencing the correct column rows 2-46
            expected_pattern = f'=AVERAGE({col_letter}2:{col_letter}46)'
            if val_upper == expected_pattern.upper():
                avg_ok_count += 1
                print(f"PASS: Component 1 — B{row_num} has correct AVERAGE formula: {val}")
            else:
                # Accept formulas referencing rows 2-46 in the correct column
                if 'AVERAGE' in val_upper and col_letter in val_upper:
                    avg_ok_count += 1
                    print(f"PASS: Component 1 — B{row_num} has AVERAGE formula: {val}")
                else:
                    print(f"FAIL: Component 1 — B{row_num} formula unexpected: {repr(val)}")

        if avg_ok_count == 5:
            print(f"PASS: Component 1 — all 5 AVERAGE formulas present in B49:B53 (0.35 pts)")
            total_score += 0.35
        elif avg_ok_count >= 3:
            partial = round(0.35 * avg_ok_count / 5, 4)
            print(f"PARTIAL: Component 1 — {avg_ok_count}/5 AVERAGE formulas present ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — only {avg_ok_count}/5 AVERAGE formulas found (0 pts)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: COUNTIF/45 formulas in C49:C53 (0.25 points)
    # Each question row must have a COUNTIF formula for >= 4 ratings divided by 45
    # FAILS on initial (all None), PASSES on golden
    try:
        countif_ok_count = 0
        for row_num, (col_letter, col_idx) in question_cols.items():
            cell = ws.cell(row=row_num, column=3)  # column C
            val = cell.value
            if val is None:
                print(f"FAIL: Component 2 — C{row_num} is empty (expected COUNTIF formula)")
                continue
            if not isinstance(val, str):
                print(f"FAIL: Component 2 — C{row_num} has non-formula value: {repr(val)}")
                continue
            val_upper = val.upper().replace(' ', '')
            # Check it's a COUNTIF formula for >= 4 divided by 45
            if 'COUNTIF' in val_upper and col_letter in val_upper and '/45' in val_upper:
                # Also check for ">=4" pattern
                if '>=4' in val or '&quot;>=4&quot;' in val_upper or '">=4"' in val:
                    countif_ok_count += 1
                    print(f"PASS: Component 2 — C{row_num} has correct COUNTIF formula: {val}")
                else:
                    # Accept if COUNTIF is present and /45 is there
                    countif_ok_count += 1
                    print(f"PASS: Component 2 — C{row_num} has COUNTIF/45 formula: {val}")
            elif 'COUNTIF' in val_upper and col_letter in val_upper:
                # Partial: has COUNTIF but missing /45
                print(f"PARTIAL: Component 2 — C{row_num} has COUNTIF but missing /45: {repr(val)}")
            else:
                print(f"FAIL: Component 2 — C{row_num} formula unexpected: {repr(val)}")

        if countif_ok_count == 5:
            print(f"PASS: Component 2 — all 5 COUNTIF/45 formulas present in C49:C53 (0.25 pts)")
            total_score += 0.25
        elif countif_ok_count >= 3:
            partial = round(0.25 * countif_ok_count / 5, 4)
            print(f"PARTIAL: Component 2 — {countif_ok_count}/5 COUNTIF formulas present ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — only {countif_ok_count}/5 COUNTIF formulas found (0 pts)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Number formats — B49:B53 formatted as 0.00, C49:C53 as percentage (0.15 points)
    # FAILS on initial (all 'General'), PASSES on golden
    try:
        fmt_ok_count = 0
        total_fmt_checks = 10

        # Check B49:B53 for 0.00 format
        for row_num in range(49, 54):
            cell = ws.cell(row=row_num, column=2)
            nf = cell.number_format
            if nf and '0.00' in nf and '%' not in nf:
                fmt_ok_count += 1
            else:
                print(f"FAIL: Component 3 — B{row_num} number format is {repr(nf)}, expected '0.00'")

        # Check C49:C53 for percentage format
        for row_num in range(49, 54):
            cell = ws.cell(row=row_num, column=3)
            nf = cell.number_format
            if nf and '%' in nf:
                fmt_ok_count += 1
            else:
                print(f"FAIL: Component 3 — C{row_num} number format is {repr(nf)}, expected percentage")

        if fmt_ok_count == total_fmt_checks:
            print(f"PASS: Component 3 — all 10 cells have correct number formats (0.15 pts)")
            total_score += 0.15
        elif fmt_ok_count >= 6:
            partial = round(0.15 * fmt_ok_count / total_fmt_checks, 4)
            print(f"PARTIAL: Component 3 — {fmt_ok_count}/{total_fmt_checks} cells have correct formats ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — only {fmt_ok_count}/{total_fmt_checks} cells have correct formats (0 pts)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Radar chart present on Survey sheet (0.15 points)
    # Initial has 0 charts, golden has 1 RadarChart — FAILS on initial, PASSES on golden
    try:
        charts = ws._charts
        radar_charts = [c for c in charts if type(c).__name__ == 'RadarChart']
        if len(radar_charts) >= 1:
            print(f"PASS: Component 4 — Radar chart found on Survey sheet (0.15 pts)")
            total_score += 0.15
        elif len(charts) >= 1:
            print(f"PARTIAL: Component 4 — Chart exists but is not a RadarChart (type: {type(charts[0]).__name__}), partial credit (0.07 pts)")
            total_score += 0.07
        else:
            print(f"FAIL: Component 4 — No chart found on Survey sheet (0 pts)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Chart title is 'Course Evaluation Radar' (0.10 points)
    # FAILS on initial (no chart, no title), PASSES on golden
    try:
        charts = ws._charts
        if charts:
            chart = charts[0]
            if chart.title is not None:
                # Extract title text using regex on the string representation
                title_str = str(chart.title)
                match = re.search(r"t='([^']+)'", title_str)
                if match:
                    title_text = match.group(1)
                    if title_text.strip() == 'Course Evaluation Radar':
                        print(f"PASS: Component 5 — Chart title is 'Course Evaluation Radar' (0.10 pts)")
                        total_score += 0.10
                    else:
                        print(f"FAIL: Component 5 — Chart title is '{title_text}', expected 'Course Evaluation Radar'")
                else:
                    print(f"FAIL: Component 5 — Could not extract chart title text from: {title_str[:100]}")
            else:
                print(f"FAIL: Component 5 — Chart has no title set")
        else:
            print(f"FAIL: Component 5 — No chart found, cannot check title")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
