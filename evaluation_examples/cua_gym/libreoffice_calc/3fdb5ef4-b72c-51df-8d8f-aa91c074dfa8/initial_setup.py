"""
Initial Setup: Course Evaluation Survey Tabulation
Task ID: calc_edu_survey_tabulation_015
Domain: libreoffice_calc

Creates a spreadsheet with:
- 'Survey' sheet with student IDs and survey responses (45 students, Q1-Q5 rated 1-5)
- Summary table placeholder (rows 48-53) with labels but NO formulas/chart yet
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_survey_tabulation_015'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # ---- Sheet: Survey ----
    ws = wb.active
    ws.title = 'Survey'

    # Headers (Row 1)
    headers = ['Student ID', 'Q1_Teaching', 'Q2_Content', 'Q3_Materials', 'Q4_Workload', 'Q5_Overall']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # 45 student responses (rows 2-46) with realistic varied ratings 1-5
    # Designed so means are: Q1~3.8, Q2~4.0, Q3~3.5, Q4~3.2, Q5~3.9
    import random
    random.seed(42)

    student_responses = [
        # (Q1, Q2, Q3, Q4, Q5)
        (4, 5, 3, 3, 4),
        (5, 5, 4, 2, 5),
        (3, 4, 3, 3, 4),
        (4, 4, 4, 4, 4),
        (5, 5, 5, 3, 5),
        (3, 3, 2, 2, 3),
        (4, 4, 3, 3, 4),
        (2, 3, 2, 2, 3),
        (5, 5, 4, 4, 5),
        (4, 4, 4, 3, 4),
        (3, 4, 3, 2, 3),
        (4, 5, 3, 3, 5),
        (5, 5, 5, 4, 5),
        (3, 3, 3, 3, 3),
        (4, 4, 4, 4, 4),
        (2, 3, 2, 2, 2),
        (5, 5, 4, 3, 5),
        (4, 4, 3, 3, 4),
        (3, 4, 2, 2, 3),
        (4, 5, 4, 3, 4),
        (5, 5, 5, 4, 5),
        (4, 4, 4, 4, 4),
        (3, 3, 3, 2, 3),
        (4, 4, 3, 3, 4),
        (5, 5, 4, 4, 5),
        (1, 2, 1, 1, 2),
        (4, 4, 4, 3, 4),
        (3, 4, 3, 3, 3),
        (5, 5, 5, 3, 5),
        (4, 4, 3, 2, 4),
        (3, 3, 2, 3, 3),
        (4, 5, 4, 4, 5),
        (5, 5, 5, 4, 5),
        (2, 3, 2, 2, 2),
        (4, 4, 3, 3, 4),
        (4, 4, 4, 3, 4),
        (3, 4, 3, 2, 3),
        (5, 5, 4, 4, 5),
        (4, 4, 4, 3, 4),
        (3, 3, 3, 3, 3),
        (4, 5, 3, 3, 4),
        (5, 5, 5, 4, 5),
        (3, 4, 3, 2, 3),
        (4, 4, 4, 3, 4),
        (4, 4, 3, 3, 4),
    ]

    for i, (q1, q2, q3, q4, q5) in enumerate(student_responses):
        row = i + 2
        ws.cell(row=row, column=1, value=f'STU{1000 + i + 1}')
        ws.cell(row=row, column=2, value=q1)
        ws.cell(row=row, column=3, value=q2)
        ws.cell(row=row, column=4, value=q3)
        ws.cell(row=row, column=5, value=q4)
        ws.cell(row=row, column=6, value=q5)

    # Row 48-53: Summary table header and labels (NO formulas yet)
    # Row 47 is blank (spacer)

    # Row 48: Summary table header row
    ws.cell(row=48, column=1, value='Question').font = Font(bold=True)
    ws.cell(row=48, column=2, value='Mean').font = Font(bold=True)
    ws.cell(row=48, column=3, value='% Satisfied (4-5)').font = Font(bold=True)

    # Rows 49-53: Question labels (B and C intentionally left empty - task requires filling them)
    question_labels = [
        'Q1_Teaching',
        'Q2_Content',
        'Q3_Materials',
        'Q4_Workload',
        'Q5_Overall',
    ]
    for i, label in enumerate(question_labels):
        row = 49 + i
        ws.cell(row=row, column=1, value=label)
        # B and C columns left empty - agent must add formulas

    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Survey')
    print(f'  Data rows: 45 (rows 2-46)')
    print(f'  Summary table: rows 48-53 (labels only, no formulas)')
    print(f'  No charts in initial file')


create_initial()
