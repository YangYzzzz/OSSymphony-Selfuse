"""
Reward Script: Transcribe expense claim photos and consolidate into expense_claims.xlsx
Task ID: osworld_multi_apps_receipt_to_calc_013
Domain: libreoffice_calc
Scoring:
  Component 1: 5 data rows populated (Employee, Date, Description, Amount, Dept) — 0.30 pts
  Component 2: Approval Required column uses nested IF formula with 500/200 thresholds  — 0.30 pts
  Component 3: Submitted Date column uses TODAY() formula                               — 0.15 pts
  Component 4: Department summary section with SUMIF/COUNTIF formulas + GRAND TOTAL     — 0.25 pts
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_receipt_to_calc_013'
FILE_PATH = f'{WORKDIR}/expense_claims.xlsx'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Scoring rubric:
      - Initial env: header row only, no data -> should score 0.0
      - Golden env: 5 data rows + IF formulas + TODAY() + department summary -> should score 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: verify the sheet 'Expense Claims' exists
    if 'Expense Claims' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Expense Claims' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Expense Claims']

    # Verify header row exists as precondition gate
    expected_headers = ['Employee', 'Claim Date', 'Description', 'Amount', 'Department',
                        'Approval Required', 'Submitted Date']
    headers_ok = True
    for col_idx, expected_header in enumerate(expected_headers, 1):
        cell_val = ws.cell(row=1, column=col_idx).value
        if cell_val != expected_header:
            headers_ok = False
            break

    if not headers_ok:
        print("CRITICAL: Header row does not match expected structure — cannot verify task")
        print("REWARD: 0.0")
        return 0.0

    # -------------------------------------------------------------------------
    # Component 1: 5 data rows populated with employee data (0.30 points)
    # The task requires transcribing 5 claim photos. Initial has only the header row.
    # Check: rows 2-6 have non-empty Employee, Claim Date, Description, Amount, Department
    # -------------------------------------------------------------------------
    try:
        populated_rows = 0
        columns_to_check = [1, 2, 3, 4, 5]  # A=Employee, B=ClaimDate, C=Desc, D=Amount, E=Dept
        for row in range(2, 7):
            row_ok = True
            for col in columns_to_check:
                val = ws.cell(row=row, column=col).value
                if val is None or str(val).strip() == '':
                    row_ok = False
                    break
            if row_ok:
                # Also verify Amount is numeric
                amount_val = ws.cell(row=row, column=4).value
                try:
                    float(amount_val)
                except (TypeError, ValueError):
                    row_ok = False
            if row_ok:
                populated_rows += 1

        if populated_rows == 5:
            print(f"PASS: Component 1 — All 5 data rows populated with Employee, Date, Description, Amount, Department (0.30 pts)")
            total_score += 0.30
        elif populated_rows >= 3:
            partial = round(0.30 * populated_rows / 5, 2)
            print(f"PARTIAL: Component 1 — {populated_rows}/5 rows fully populated ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {populated_rows}/5 data rows fully populated (expected 5)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Approval Required column (F2:F6) uses nested IF formula
    # with thresholds >500 -> Director, >200 -> Manager, else Self-Approve (0.30 points)
    # Initial has no data rows, so this trivially fails on initial.
    # -------------------------------------------------------------------------
    try:
        approval_formula_count = 0
        for row in range(2, 7):
            cell_val = ws.cell(row=row, column=6).value
            if cell_val is None:
                continue
            val_str = str(cell_val).upper().replace(' ', '')
            # Check for nested IF with 500 and 200 thresholds and expected approval levels
            if (val_str.startswith('=IF') and
                    '>500' in val_str and '>200' in val_str and
                    'DIRECTOR' in val_str and 'MANAGER' in val_str and
                    'SELF-APPROVE' in val_str):
                approval_formula_count += 1
            elif (val_str.startswith('=IF') and
                    '500' in val_str and '200' in val_str and
                    'DIRECTOR' in val_str and 'MANAGER' in val_str):
                # Partial match - has IF formula with thresholds but may differ slightly
                approval_formula_count += 0.5

        if approval_formula_count >= 5:
            print(f"PASS: Component 2 — All 5 rows have nested IF formula with 500/200 thresholds and correct approval levels (0.30 pts)")
            total_score += 0.30
        elif approval_formula_count >= 3:
            partial = round(0.30 * approval_formula_count / 5, 2)
            print(f"PARTIAL: Component 2 — {approval_formula_count}/5 rows have correct Approval Required formula ({partial} pts)")
            total_score += partial
        else:
            # Check if any approval values are present but hardcoded (not formula)
            # This means task was attempted but incorrectly
            hardcoded_count = 0
            for row in range(2, 7):
                cell_val = ws.cell(row=row, column=6).value
                if cell_val in ('Director', 'Manager', 'Self-Approve'):
                    hardcoded_count += 1
            if hardcoded_count > 0:
                print(f"FAIL: Component 2 — {hardcoded_count} rows have hardcoded approval values instead of IF formulas. Expected formula like =IF(D2>500,\"Director\",IF(D2>200,\"Manager\",\"Self-Approve\"))")
            else:
                print(f"FAIL: Component 2 — No valid nested IF approval formulas found (found {approval_formula_count}/5)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Submitted Date column (G2:G6) uses =TODAY() formula (0.15 points)
    # Initial has no data rows, so this trivially fails on initial.
    # -------------------------------------------------------------------------
    try:
        today_formula_count = 0
        for row in range(2, 7):
            cell_val = ws.cell(row=row, column=7).value
            if cell_val is None:
                continue
            val_str = str(cell_val).upper().replace(' ', '')
            if val_str == '=TODAY()':
                today_formula_count += 1

        if today_formula_count == 5:
            print(f"PASS: Component 3 — All 5 rows have =TODAY() formula in Submitted Date column (0.15 pts)")
            total_score += 0.15
        elif today_formula_count >= 3:
            partial = round(0.15 * today_formula_count / 5, 2)
            print(f"PARTIAL: Component 3 — {today_formula_count}/5 rows have =TODAY() formula ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {today_formula_count}/5 rows have =TODAY() formula in Submitted Date column")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: Department summary section with SUMIF/COUNTIF formulas (0.25 points)
    # Expected: rows 8-15 with a header row, dept rows with SUMIF/COUNTIF, and GRAND TOTAL
    # Initial has no summary section, so trivially fails on initial.
    # -------------------------------------------------------------------------
    try:
        summary_score = 0.0

        # Sub-check 4a: Summary header label at A8
        a8_val = ws.cell(row=8, column=1).value
        if a8_val and 'summary' in str(a8_val).lower():
            summary_score += 0.05
            print(f"  PASS: Sub-check 4a — Summary header label found at A8: '{a8_val}'")
        else:
            print(f"  FAIL: Sub-check 4a — No summary label found at A8 (found: {repr(a8_val)})")

        # Sub-check 4b: Department summary rows (rows 10-14) have SUMIF and COUNTIF formulas
        sumif_count = 0
        countif_count = 0
        for row in range(10, 15):
            b_val = ws.cell(row=row, column=2).value
            c_val = ws.cell(row=row, column=3).value
            if b_val and isinstance(b_val, str) and 'SUMIF' in b_val.upper():
                sumif_count += 1
            if c_val and isinstance(c_val, str) and 'COUNTIF' in c_val.upper():
                countif_count += 1

        if sumif_count >= 5:
            summary_score += 0.10
            print(f"  PASS: Sub-check 4b — All 5 dept rows have SUMIF formula in column B")
        elif sumif_count >= 3:
            partial = round(0.10 * sumif_count / 5, 2)
            summary_score += partial
            print(f"  PARTIAL: Sub-check 4b — {sumif_count}/5 dept rows have SUMIF formula ({partial} pts)")
        else:
            print(f"  FAIL: Sub-check 4b — Only {sumif_count}/5 dept rows have SUMIF formula in column B")

        if countif_count >= 5:
            summary_score += 0.05
            print(f"  PASS: Sub-check 4c — All 5 dept rows have COUNTIF formula in column C")
        elif countif_count >= 3:
            partial = round(0.05 * countif_count / 5, 2)
            summary_score += partial
            print(f"  PARTIAL: Sub-check 4c — {countif_count}/5 dept rows have COUNTIF formula ({partial} pts)")
        else:
            print(f"  FAIL: Sub-check 4c — Only {countif_count}/5 dept rows have COUNTIF formula in column C")

        # Sub-check 4d: GRAND TOTAL row (row 15) with SUM formulas
        a15_val = ws.cell(row=15, column=1).value
        b15_val = ws.cell(row=15, column=2).value
        c15_val = ws.cell(row=15, column=3).value
        has_grand_total_label = a15_val and 'total' in str(a15_val).lower()
        has_sum_b = b15_val and isinstance(b15_val, str) and 'SUM' in b15_val.upper()
        has_sum_c = c15_val and isinstance(c15_val, str) and 'SUM' in c15_val.upper()

        if has_grand_total_label and has_sum_b and has_sum_c:
            summary_score += 0.05
            print(f"  PASS: Sub-check 4d — GRAND TOTAL row found at row 15 with SUM formulas")
        else:
            print(f"  FAIL: Sub-check 4d — GRAND TOTAL row missing or incomplete: label={repr(a15_val)}, B15={repr(b15_val)}, C15={repr(c15_val)}")

        if summary_score >= 0.25:
            print(f"PASS: Component 4 — Department summary section complete (0.25 pts)")
            total_score += 0.25
        elif summary_score > 0:
            total_score += min(summary_score, 0.25)
            print(f"PARTIAL: Component 4 — Department summary section partially complete ({summary_score:.2f} pts)")
        else:
            print(f"FAIL: Component 4 — Department summary section not found or empty")

    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


if not os.path.exists(FILE_PATH):
    print(f"File not found: {FILE_PATH}")
    print("REWARD: 0.0")
else:
    verify_task(FILE_PATH)
