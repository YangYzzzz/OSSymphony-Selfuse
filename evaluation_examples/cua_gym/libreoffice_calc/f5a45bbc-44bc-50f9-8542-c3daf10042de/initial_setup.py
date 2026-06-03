"""
Initial Setup: Expense Claims Spreadsheet with Handwritten Receipt Photos
Task ID: osworld_multi_apps_receipt_to_calc_013
Domain: libreoffice_calc

Creates:
  - /home/user/expense_claims.xlsx — empty spreadsheet with headers only (no data rows)
  - 5 JPEG images on the desktop representing handwritten expense claim forms
  - Opens expense_claims.xlsx in LibreOffice Calc
"""

import os
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# PIL for creating realistic-looking handwritten receipt images
try:
    from PIL import Image, ImageDraw, ImageFont
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_receipt_to_calc_013'
DESKTOP = '/home/user/Desktop'
OUTPUT = f'{WORKDIR}/expense_claims.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_receipt_image_pil(filepath, claim_data):
    """Create a realistic-looking handwritten expense claim photo using PIL."""
    # Create a white paper-like background (A4 proportion roughly)
    width, height = 800, 600
    img = Image.new('RGB', (width, height), color=(252, 250, 245))
    draw = ImageDraw.Draw(img)

    # Draw a light yellow/cream background to mimic paper
    for y in range(height):
        alpha = y / height * 10
        r = int(252 - alpha)
        g = int(250 - alpha)
        b = int(240 - alpha)
        draw.line([(0, y), (width, y)], fill=(r, g, b))

    # Draw border
    draw.rectangle([20, 20, width-20, height-20], outline=(100, 100, 100), width=2)

    # Title
    draw.rectangle([20, 20, width-20, 70], fill=(220, 230, 255), outline=(100, 100, 100), width=1)
    draw.text((width//2 - 140, 35), "EMPLOYEE EXPENSE CLAIM FORM", fill=(20, 20, 80),
              font=None)

    # Form fields
    y_pos = 90
    line_height = 45
    label_x = 40
    value_x = 200

    fields = [
        ("Employee Name:", claim_data['employee']),
        ("Department:", claim_data['department']),
        ("Claim Date:", claim_data['claim_date']),
        ("Description:", claim_data['description']),
        ("Amount ($):", f"${claim_data['amount']:.2f}"),
        ("Signature:", claim_data['signature']),
    ]

    for label, value in fields:
        # Draw light line background alternating
        draw.text((label_x, y_pos), label, fill=(80, 80, 80), font=None)
        draw.text((value_x, y_pos), value, fill=(10, 10, 10), font=None)
        # Underline for value
        text_width = len(value) * 7
        draw.line([(value_x, y_pos + 16), (value_x + max(text_width, 200), y_pos + 16)],
                  fill=(150, 150, 150), width=1)
        y_pos += line_height

    # Add some "handwritten" feel with a slight slant note
    draw.text((label_x, y_pos + 10),
              "Please attach receipts and submit to Finance dept.",
              fill=(120, 100, 80), font=None)

    # Add a stamp-like circle for authenticity
    draw.ellipse([width-120, height-120, width-40, height-40],
                 outline=(180, 50, 50), width=3)
    draw.text((width-105, height-90), "PENDING", fill=(180, 50, 50), font=None)
    draw.text((width-105, height-75), "REVIEW", fill=(180, 50, 50), font=None)

    img.save(filepath, 'JPEG', quality=85)


def create_receipt_image_fallback(filepath, claim_data):
    """Create receipt image using basic drawing without PIL fonts."""
    import struct
    import zlib

    # Create a simple PPM then convert or just write a tiny valid JPEG
    # Use PIL basic mode
    width, height = 640, 480
    img = Image.new('RGB', (width, height), color=(255, 253, 240))
    draw = ImageDraw.Draw(img)

    # Header band
    draw.rectangle([0, 0, width, 60], fill=(200, 215, 250))
    draw.text((10, 15), "EXPENSE CLAIM FORM", fill=(30, 30, 120))
    draw.text((10, 35), f"Ref: {claim_data['employee'][:3].upper()}-{claim_data['claim_date'][-5:]}", fill=(80, 80, 120))

    y = 80
    for label, value in [
        ("Name:", claim_data['employee']),
        ("Dept:", claim_data['department']),
        ("Date:", claim_data['claim_date']),
        ("Item:", claim_data['description']),
        ("Amt: $", f"{claim_data['amount']:.2f}"),
    ]:
        draw.text((20, y), label, fill=(100, 80, 60))
        draw.text((100, y), str(value), fill=(20, 20, 20))
        draw.line([(100, y+18), (400, y+18)], fill=(180, 180, 180))
        y += 45

    draw.text((20, y+10), f"Signed: {claim_data['signature']}", fill=(60, 60, 60))
    draw.text((20, y+30), "Submit to Finance for approval.", fill=(120, 100, 80))

    img.save(filepath, 'JPEG', quality=88)


def create_initial():
    os.makedirs(DESKTOP, exist_ok=True)

    # ---------------------------------------------------------
    # 1. Create expense_claims.xlsx with headers only (no data)
    # ---------------------------------------------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expense Claims"

    # Column headers
    headers = [
        "Employee",
        "Claim Date",
        "Description",
        "Amount",
        "Department",
        "Approval Required",
        "Submitted Date"
    ]

    # Header styling
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Set column widths
    col_widths = {
        "A": 22,   # Employee
        "B": 14,   # Claim Date
        "C": 32,   # Description
        "D": 12,   # Amount
        "E": 18,   # Department
        "F": 20,   # Approval Required
        "G": 16,   # Submitted Date
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Set header row height
    ws.row_dimensions[1].height = 30

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Number format for Amount column (no data yet, but set column format)
    # We just set the header; data rows will be added by agent

    wb.save(OUTPUT)
    print(f"Initial file created: {OUTPUT}")

    # ---------------------------------------------------------
    # 2. Create 5 JPEG photos on Desktop of expense claim forms
    # ---------------------------------------------------------
    claims_for_photos = [
        {
            "employee": "Rachel Kim",
            "department": "Marketing",
            "claim_date": "2025-11-03",
            "description": "Client dinner - Q4 planning",
            "amount": 347.50,
            "signature": "R. Kim",
        },
        {
            "employee": "James Okafor",
            "department": "Engineering",
            "claim_date": "2025-11-07",
            "description": "Conference travel - DevConf 2025",
            "amount": 782.00,
            "signature": "J. Okafor",
        },
        {
            "employee": "Priya Sharma",
            "department": "Sales",
            "claim_date": "2025-11-10",
            "description": "Office supplies restock",
            "amount": 88.25,
            "signature": "P. Sharma",
        },
        {
            "employee": "Daniel Torres",
            "department": "Operations",
            "claim_date": "2025-11-12",
            "description": "Team lunch - project kickoff",
            "amount": 245.00,
            "signature": "D. Torres",
        },
        {
            "employee": "Mei-Ling Zhang",
            "department": "Finance",
            "claim_date": "2025-11-15",
            "description": "Software subscription renewal",
            "amount": 599.99,
            "signature": "M. Zhang",
        },
    ]

    if HAS_PIL:
        for i, claim in enumerate(claims_for_photos, 1):
            photo_path = f"{DESKTOP}/expense_claim_photo_{i:02d}.jpg"
            try:
                create_receipt_image_pil(photo_path, claim)
            except Exception:
                create_receipt_image_fallback(photo_path, claim)
            print(f"Created photo: {photo_path}")
    else:
        # Fallback: create minimal placeholder JPEG files using raw bytes
        # Minimal valid JPEG (1x1 pixel white)
        minimal_jpeg = bytes([
            0xFF, 0xD8, 0xFF, 0xE0, 0x00, 0x10, 0x4A, 0x46, 0x49, 0x46, 0x00, 0x01,
            0x01, 0x00, 0x00, 0x01, 0x00, 0x01, 0x00, 0x00, 0xFF, 0xDB, 0x00, 0x43,
            0x00, 0x08, 0x06, 0x06, 0x07, 0x06, 0x05, 0x08, 0x07, 0x07, 0x07, 0x09,
            0x09, 0x08, 0x0A, 0x0C, 0x14, 0x0D, 0x0C, 0x0B, 0x0B, 0x0C, 0x19, 0x12,
            0x13, 0x0F, 0x14, 0x1D, 0x1A, 0x1F, 0x1E, 0x1D, 0x1A, 0x1C, 0x1C, 0x20,
            0x24, 0x2E, 0x27, 0x20, 0x22, 0x2C, 0x23, 0x1C, 0x1C, 0x28, 0x37, 0x29,
            0x2C, 0x30, 0x31, 0x34, 0x34, 0x34, 0x1F, 0x27, 0x39, 0x3D, 0x38, 0x32,
            0x3C, 0x2E, 0x33, 0x34, 0x32, 0xFF, 0xC0, 0x00, 0x0B, 0x08, 0x00, 0x01,
            0x00, 0x01, 0x01, 0x01, 0x11, 0x00, 0xFF, 0xC4, 0x00, 0x1F, 0x00, 0x00,
            0x01, 0x05, 0x01, 0x01, 0x01, 0x01, 0x01, 0x01, 0x00, 0x00, 0x00, 0x00,
            0x00, 0x00, 0x00, 0x00, 0x01, 0x02, 0x03, 0x04, 0x05, 0x06, 0x07, 0x08,
            0x09, 0x0A, 0x0B, 0xFF, 0xC4, 0x00, 0xB5, 0x10, 0x00, 0x02, 0x01, 0x03,
            0x03, 0x02, 0x04, 0x03, 0x05, 0x05, 0x04, 0x04, 0x00, 0x00, 0x01, 0x7D,
            0xFF, 0xDA, 0x00, 0x08, 0x01, 0x01, 0x00, 0x00, 0x3F, 0x00, 0xFB, 0xD2,
            0x8A, 0x28, 0x03, 0xFF, 0xD9,
        ])
        for i, claim in enumerate(claims_for_photos, 1):
            photo_path = f"{DESKTOP}/expense_claim_photo_{i:02d}.jpg"
            with open(photo_path, 'wb') as f:
                f.write(minimal_jpeg)
            print(f"Created placeholder photo: {photo_path}")

    # ---------------------------------------------------------
    # 3. GUI-ready startup: open expense_claims.xlsx in Calc
    # ---------------------------------------------------------
    # Also open Files/Nautilus to show desktop photos
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    # Open the desktop folder so photos are visible
    launch_gui(f'nautilus "{DESKTOP}"', delay_sec=1.5)
    print("GUI_READY: launched LibreOffice Calc and Nautilus with DISPLAY=:0")


create_initial()
