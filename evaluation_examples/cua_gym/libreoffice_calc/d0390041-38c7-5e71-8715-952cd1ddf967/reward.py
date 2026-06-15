"""
FINAL REWARD SCRIPT - SUCCESS
Task: Please create a Pivot Table in Sheet2 that summarizes the total orders for each region with region names as the column headers.
Generated: 2025-11-24 07:48:12
Status: success
Model: o3
Total Steps: 3
"""

import openpyxl
import math
import re


def verify_pivot_table(file_path: str) -> float:
    """Verify that Sheet2 contains a pivot-style summary of total orders per
    region, with region names as column headers and totals underneath.

    Scoring (progressive):
        0.4  – header row contains ALL region names
        0.6  – a totals row matches expected sums **or** contains correct
                 SUMIF-style formulas referencing Sheet1
        1.0  – both conditions satisfied (0.4 + 0.6)
    """
    max_score = 1.0
    score = 0.0

    print(f"Verifying file: {file_path}")

    # ------------------------------------------------------------------
    # 1. Load workbook twice:  formulas view & values-only view
    # ------------------------------------------------------------------
    try:
        wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
        wb_values = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"✗ Failed to load workbook: {e}")
        return 0.0

    # ------------------------------------------------------------------
    # 2. Compute expected totals from Sheet1
    # ------------------------------------------------------------------
    if "Sheet1" not in wb_formulas.sheetnames:
        print("✗ Sheet1 not found in workbook")
        return 0.0

    sh1 = wb_formulas["Sheet1"]
    header = [str(c.value).strip() if c.value is not None else "" for c in sh1[1]]
    try:
        region_idx = header.index("Region")
        orders_idx = header.index("Orders")
    except ValueError:
        print("✗ 'Region' or 'Orders' column missing in Sheet1 header")
        return 0.0

    expected_totals = {}
    for row in sh1.iter_rows(min_row=2, values_only=True):
        region = row[region_idx]
        orders = row[orders_idx]
        if region is None or orders is None:
            continue
        try:
            orders_val = float(orders)
        except (ValueError, TypeError):
            continue
        expected_totals[region] = expected_totals.get(region, 0) + orders_val

    print("Expected totals based on Sheet1:", expected_totals)

    # ------------------------------------------------------------------
    # 3. Analyse Sheet2 (pivot summary)
    # ------------------------------------------------------------------
    if "Sheet2" not in wb_formulas.sheetnames:
        print("✗ Sheet2 not found in workbook")
        return 0.0

    sh2_f = wb_formulas["Sheet2"]  # formulas view
    sh2_v = wb_values["Sheet2"]    # values view

    max_row, max_col = sh2_f.max_row, sh2_f.max_column
    if max_row == 0 or max_col == 0:
        print("✗ Sheet2 is empty")
        return 0.0

    # Build matrices of formulas & evaluated values for easy access
    matrix_f = [
        [sh2_f.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        for r in range(1, max_row + 1)
    ]
    matrix_v = [
        [sh2_v.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        for r in range(1, max_row + 1)
    ]

    # ------------------------------------------------------------------
    # 3a. Locate header row containing ALL region names
    # ------------------------------------------------------------------
    regions = list(expected_totals.keys())
    header_row_idx = None
    region_to_col = {}

    for r_idx, row in enumerate(matrix_f):
        present = {
            str(val).strip().lower(): c_idx
            for c_idx, val in enumerate(row)
            if isinstance(val, str)
        }
        if all(reg.lower() in present for reg in regions):
            header_row_idx = r_idx
            region_to_col = {reg: present[reg.lower()] for reg in regions}
            break

    if header_row_idx is None:
        print("✗ Could not locate header row containing all region names")
    else:
        print(f"✓ Header row found at Excel row {header_row_idx + 1}")
        score += 0.4  # header present → 0.4 pts

    # ------------------------------------------------------------------
    # 3b. Verify totals row (either numeric results or correct SUMIF formula)
    # ------------------------------------------------------------------
    totals_verified = False

    if header_row_idx is not None:
        for r_idx in range(header_row_idx + 1, max_row):
            row_f = matrix_f[r_idx]
            row_v = matrix_v[r_idx]

            # Heuristic: first cell text contains the word "total"
            first_cell = row_f[0] if row_f else None
            if not (isinstance(first_cell, str) and "total" in first_cell.lower()):
                continue

            all_match = True
            for reg, col in region_to_col.items():
                expected = expected_totals[reg]
                val_value = row_v[col]
                val_formula = row_f[col]
                numeric_ok = (
                    isinstance(val_value, (int, float)) and
                    math.isclose(float(val_value), expected, rel_tol=0.01, abs_tol=1)
                )

                if numeric_ok:
                    continue  # value matches expected

                # If not numeric (e.g., not calculated), verify formula pattern
                if isinstance(val_formula, str) and val_formula.startswith("="):
                    col_letter = openpyxl.utils.get_column_letter(col + 1)
                    header_row_excel = header_row_idx + 1  # convert 0-based → 1-based
                    pattern = rf"=SUMIF\(Sheet1!\$?[A-Z]+:\$?[A-Z]+,\s*{col_letter}\$?{header_row_excel},\s*Sheet1!\$?[A-Z]+:\$?[A-Z]+\)"
                    if re.match(pattern, val_formula, re.IGNORECASE):
                        continue  # formula appears correct

                # If neither numeric nor acceptable formula → fail
                print(
                    f"✗ Region '{reg}' column mismatch in row {r_idx + 1}: "
                    f"value={val_value}, formula={val_formula}"
                )
                all_match = False
                break

            if all_match:
                totals_verified = True
                print(f"✓ Totals row verified at Excel row {r_idx + 1}")
                break

        if not totals_verified:
            print("✗ Unable to verify a correct totals row in Sheet2")
        else:
            score += 0.6  # correct totals → 0.6 pts

    # ------------------------------------------------------------------
    # 4. Final score
    # ------------------------------------------------------------------
    final_score = round(min(score, max_score), 2)
    print(f"Total score awarded: {final_score}")
    return final_score


if __name__ == "__main__":
    # Default path inside VM
    FILE_PATH = "/home/user/please_create_a_pivot_table_in_sheet2_that_summarizes_the_total_orders_for_each_region_with_region_n.xlsx"
    reward = verify_pivot_table(FILE_PATH)
    print(f"REWARD: {reward}")
