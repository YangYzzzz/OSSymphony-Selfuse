"""
Initial Setup: Build a dynamic employee cost center report
Task ID: calc_hr_089
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_089'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()

    # --- Sheet 'Costs' ---
    ws_costs = wb.active
    ws_costs.title = 'Costs'

    headers = ['Employee', 'Cost Center', 'Project', 'Month', 'Hours', 'Rate', 'Cost']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")

    for col, h in enumerate(headers, 1):
        cell = ws_costs.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.font = Font(bold=True, size=11, color="FFFFFF")

    # Data pools
    cost_centers = ['Engineering', 'Marketing', 'Sales', 'Operations', 'Finance']
    projects = ['Alpha', 'Beta', 'Gamma', 'Delta', 'Epsilon', 'Zeta', 'Eta', 'Theta']
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun']

    employees = [
        'Sarah Chen', 'Marcus Johnson', 'Priya Patel', 'James O\'Brien',
        'Yuki Tanaka', 'Elena Rodriguez', 'David Kim', 'Fatima Al-Hassan',
        'Robert Fischer', 'Mei Lin Wang', 'Carlos Mendez', 'Anna Kowalski',
        'Tariq Ahmed', 'Lisa Nguyen', 'Michael Brown', 'Sophia Larsen',
        'Omar Diallo', 'Rachel Green', 'Kenji Yamamoto', 'Isabella Rossi',
        'Thomas Mueller', 'Aisha Okafor', 'Daniel Park', 'Maria Santos',
        'Henrik Johansson'
    ]

    rates = [45.00, 55.00, 65.00, 75.00, 85.00, 95.00, 105.00, 115.00, 125.00]

    # Generate 100 rows of realistic data
    for r in range(2, 102):
        emp = random.choice(employees)
        cc = random.choice(cost_centers)
        proj = random.choice(projects)
        month = random.choice(months)
        hours = random.randint(4, 40)
        rate = random.choice(rates)
        cost = round(hours * rate, 2)

        ws_costs.cell(row=r, column=1, value=emp)
        ws_costs.cell(row=r, column=2, value=cc)
        ws_costs.cell(row=r, column=3, value=proj)
        ws_costs.cell(row=r, column=4, value=month)
        ws_costs.cell(row=r, column=5, value=hours)
        ws_costs.cell(row=r, column=6, value=rate)
        ws_costs.cell(row=r, column=7, value=cost)

    # Format number columns
    for r in range(2, 102):
        ws_costs.cell(row=r, column=6).number_format = '$#,##0.00'
        ws_costs.cell(row=r, column=7).number_format = '$#,##0.00'

    # Set column widths
    ws_costs.column_dimensions['A'].width = 22
    ws_costs.column_dimensions['B'].width = 16
    ws_costs.column_dimensions['C'].width = 14
    ws_costs.column_dimensions['D'].width = 10
    ws_costs.column_dimensions['E'].width = 10
    ws_costs.column_dimensions['F'].width = 10
    ws_costs.column_dimensions['G'].width = 12

    ws_costs.freeze_panes = 'A2'

    # --- Sheet 'Report' (EMPTY - agent must build it) ---
    ws_report = wb.create_sheet('Report')
    # Leave completely empty - the task is to build the report

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
