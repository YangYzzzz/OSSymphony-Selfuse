"""
Reward Script: Build a dynamic employee cost center report with SUMPRODUCT formulas
Task ID: calc_hr_089
Domain: libreoffice_calc
Scoring:
  Component 1 (0.15): Report sheet has correct headers (Cost Center + 6 months + Total)
  Component 2 (0.15): Report has 5 cost center row labels
  Component 3 (0.35): SUMPRODUCT formulas in the 30 data cells (5 cost centers x 6 months)
  Component 4 (0.15): Row totals (SUM formulas in column H for each cost center)
  Component 5 (0.20): Column totals row with SUM formulas + grand total
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_089'

# Expected structure from task description
EXPECTED_COST_CENTERS = ['Engineering', 'Marketing', 'Sales', 'Operations', 'Finance']
EXPECTED_MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun']


def persist_app_state(domain):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(1.0)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def is_sumproduct_formula(value, cost_center, month):
    """Check if a cell value is a valid SUMPRODUCT formula filtering by cost_center and month."""
    if not isinstance(value, str):
        return False
    val_upper = value.upper().replace(" ", "")
    # Must start with =SUMPRODUCT
    if not val_upper.startswith("=SUMPRODUCT("):
        return False
    # Must reference the cost center name and month name
    if cost_center.upper() not in val_upper and f'"{cost_center.upper()}"' not in value.upper():
        # Check case-insensitive
        if cost_center.lower() not in value.lower():
            return False
    if month.upper() not in val_upper and f'"{month.upper()}"' not in value.upper():
        if month.lower() not in value.lower():
            return False
    return True


def is_sum_formula(value):
    """Check if a cell contains a SUM formula."""
    if not isinstance(value, str):
        return False
    val_upper = value.upper().replace(" ", "")
    return val_upper.startswith("=SUM(")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Report sheet must exist
    if 'Report' not in wb.sheetnames:
        print("FAIL: 'Report' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Report']

    # Component 1: Report headers (0.15 points)
    # The golden has: A1='Cost Center', B1='Jan', C1='Feb', D1='Mar', E1='Apr', F1='May', G1='Jun', H1='Total'
    # We check that the header row has a label column and 6 month columns + total
    try:
        header_row = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            if v is not None:
                header_row.append(str(v).strip())

        # Check months are present in headers
        months_found = 0
        for month in EXPECTED_MONTHS:
            if any(month.lower() in h.lower() for h in header_row):
                months_found += 1

        has_total_header = any('total' in h.lower() for h in header_row)
        has_label_header = len(header_row) >= 8  # At least label + 6 months + total

        if months_found == 6 and has_total_header and has_label_header:
            print(f"PASS: Component 1 -- Headers correct: {header_row} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 -- Expected 6 months + Total in headers, found: {header_row} (months={months_found}, total={has_total_header})")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Cost center row labels (0.15 points)
    # 5 cost centers should appear as row labels in column A (rows 2-6 or similar)
    try:
        found_centers = set()
        center_rows = {}  # map cost_center -> row number
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if val is not None:
                val_str = str(val).strip()
                for cc in EXPECTED_COST_CENTERS:
                    if val_str.lower() == cc.lower():
                        found_centers.add(cc)
                        center_rows[cc] = r

        if len(found_centers) == 5:
            print(f"PASS: Component 2 -- All 5 cost centers found as row labels: {sorted(found_centers)} (0.15 pts)")
            total_score += 0.15
        elif len(found_centers) >= 3:
            partial = 0.15 * len(found_centers) / 5
            print(f"PARTIAL: Component 2 -- {len(found_centers)}/5 cost centers found: {sorted(found_centers)} ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- Only {len(found_centers)}/5 cost centers found: {sorted(found_centers)}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: SUMPRODUCT formulas in data cells (0.35 points)
    # Each of the 30 cells (5 cost centers x 6 months) should have a SUMPRODUCT formula
    try:
        # Determine month-to-column mapping from headers
        month_cols = {}
        for c in range(2, ws.max_column + 1):
            hval = ws.cell(row=1, column=c).value
            if hval is not None:
                hstr = str(hval).strip()
                for month in EXPECTED_MONTHS:
                    if hstr.lower() == month.lower():
                        month_cols[month] = c

        sumproduct_count = 0
        total_cells = 0
        for cc in EXPECTED_COST_CENTERS:
            if cc not in center_rows:
                continue
            row = center_rows[cc]
            for month in EXPECTED_MONTHS:
                if month not in month_cols:
                    continue
                col = month_cols[month]
                total_cells += 1
                cell_val = ws.cell(row=row, column=col).value
                if is_sumproduct_formula(cell_val, cc, month):
                    sumproduct_count += 1

        if total_cells > 0:
            if sumproduct_count == 30:
                print(f"PASS: Component 3 -- All 30 SUMPRODUCT formulas present (0.35 pts)")
                total_score += 0.35
            elif sumproduct_count > 0:
                partial_pts = 0.35 * sumproduct_count / 30
                print(f"PARTIAL: Component 3 -- {sumproduct_count}/30 SUMPRODUCT formulas found ({partial_pts:.3f} pts)")
                total_score += partial_pts
            else:
                print(f"FAIL: Component 3 -- No SUMPRODUCT formulas found in data cells")
        else:
            print(f"FAIL: Component 3 -- Could not locate data cells (centers={len(center_rows)}, month_cols={len(month_cols)})")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Row totals - SUM formulas in the last column for each cost center (0.15 points)
    try:
        # Find the Total column
        total_col = None
        for c in range(2, ws.max_column + 1):
            hval = ws.cell(row=1, column=c).value
            if hval is not None and 'total' in str(hval).strip().lower():
                total_col = c
                break

        if total_col is None:
            # Try the column after the last month
            if month_cols:
                total_col = max(month_cols.values()) + 1

        row_sum_count = 0
        if total_col:
            for cc in EXPECTED_COST_CENTERS:
                if cc not in center_rows:
                    continue
                row = center_rows[cc]
                cell_val = ws.cell(row=row, column=total_col).value
                if is_sum_formula(cell_val):
                    row_sum_count += 1

        if row_sum_count == 5:
            print(f"PASS: Component 4 -- All 5 row total SUM formulas found in col {total_col} (0.15 pts)")
            total_score += 0.15
        elif row_sum_count > 0:
            partial = 0.15 * row_sum_count / 5
            print(f"PARTIAL: Component 4 -- {row_sum_count}/5 row total SUM formulas ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 -- No row total SUM formulas found")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # Component 5: Column totals row + grand total (0.20 points)
    # There should be a totals row at the bottom with SUM formulas for each month column and grand total
    try:
        # Find the totals row (look for 'Total' label in column A after the cost centers)
        totals_row = None
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if val is not None and 'total' in str(val).strip().lower():
                totals_row = r

        col_sum_count = 0
        if totals_row:
            # Check SUM formulas for each month column
            for month in EXPECTED_MONTHS:
                if month not in month_cols:
                    continue
                col = month_cols[month]
                cell_val = ws.cell(row=totals_row, column=col).value
                if is_sum_formula(cell_val):
                    col_sum_count += 1

        # Score: 6 column sums (0.12 pts)
        if col_sum_count == 6:
            print(f"PASS: Component 5a -- All 6 column total SUM formulas found in row {totals_row} (0.12 pts)")
            total_score += 0.12
        elif col_sum_count > 0:
            partial_5a = 0.12 * col_sum_count / 6
            print(f"PARTIAL: Component 5a -- {col_sum_count}/6 column total SUM formulas ({partial_5a:.3f} pts)")
            total_score += partial_5a
        else:
            print(f"FAIL: Component 5a -- No column total SUM formulas found (totals_row={totals_row})")

        # Grand total (0.08 pts)
        if totals_row and total_col and is_sum_formula(ws.cell(row=totals_row, column=total_col).value):
            print(f"PASS: Component 5b -- Grand total SUM formula found at row {totals_row}, col {total_col} (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 5b -- Grand total SUM formula not found")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
