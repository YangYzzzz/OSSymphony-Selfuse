"""
Reward Script: Create Pareto chart from defect data
Task ID: calc_gcp_068
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): At least one chart exists on the DefectAnalysis sheet
  Component 2 (0.20): A BarChart sub-chart references Count data (column B)
  Component 3 (0.20): A LineChart sub-chart references CumulativePercent data (column C)
  Component 4 (0.15): Column D contains 80% reference line data (all 0.8)
  Component 5 (0.15): Chart title contains 'Pareto'
"""

import os

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_068'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def get_title_text(title_obj):
    """Extract plain text from an openpyxl chart Title object."""
    if title_obj is None:
        return None
    try:
        for p in title_obj.tx.rich.paragraphs:
            for r in p.r:
                if r.t:
                    return r.t
    except Exception:
        pass
    return None


def verify_task(file_path):
    """
    Verify Pareto chart creation with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl
    from openpyxl.chart import BarChart, LineChart

    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: DefectAnalysis sheet must exist
    if 'DefectAnalysis' not in wb.sheetnames:
        print("FAIL: Sheet 'DefectAnalysis' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['DefectAnalysis']

    # Component 1: At least one chart exists on the sheet (0.30 points)
    # Initial file has 0 charts; golden has >= 1
    try:
        chart_count = len(ws._charts)
        if chart_count >= 1:
            print(f"PASS: Component 1 — {chart_count} chart(s) found on DefectAnalysis (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 — No charts found on DefectAnalysis sheet")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # If no charts, remaining chart-based checks will fail; skip gracefully
    if len(ws._charts) == 0:
        # Still check Component 4 (column D data) independently
        pass
    else:
        chart = ws._charts[0]

        # Collect sub-chart types and their data references
        sub_charts = []
        if hasattr(chart, '_charts') and chart._charts:
            for sc in chart._charts:
                sc_type = type(sc).__name__
                refs = []
                for s in sc.series:
                    try:
                        ref = s.val.numRef.f if s.val and hasattr(s.val, 'numRef') and s.val.numRef else None
                    except Exception:
                        ref = None
                    refs.append(ref)
                sub_charts.append((sc_type, refs))
        else:
            # Single chart (not combination) — treat the main chart as one sub-chart
            sc_type = type(chart).__name__
            refs = []
            for s in chart.series:
                try:
                    ref = s.val.numRef.f if s.val and hasattr(s.val, 'numRef') and s.val.numRef else None
                except Exception:
                    ref = None
                refs.append(ref)
            sub_charts.append((sc_type, refs))

        print(f"  DEBUG: Sub-charts detected: {sub_charts}")

        # Component 2: A BarChart references Count data in column B (0.20 points)
        try:
            has_bar_with_count = False
            for sc_type, refs in sub_charts:
                if sc_type == 'BarChart':
                    for ref in refs:
                        if ref and ('$B$' in ref or '!B' in ref.upper() or 'B2' in ref.upper()):
                            has_bar_with_count = True
                            break
            if has_bar_with_count:
                print(f"PASS: Component 2 — BarChart sub-chart references Count (column B) data (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 2 — No BarChart sub-chart referencing column B (Count) found")
        except Exception as e:
            print(f"ERROR: Component 2 — {e}")

        # Component 3: A LineChart references CumulativePercent data in column C (0.20 points)
        try:
            has_line_with_cumulative = False
            for sc_type, refs in sub_charts:
                if sc_type == 'LineChart':
                    for ref in refs:
                        if ref and ('$C$' in ref or '!C' in ref.upper() or 'C2' in ref.upper()):
                            has_line_with_cumulative = True
                            break
            if has_line_with_cumulative:
                print(f"PASS: Component 3 — LineChart sub-chart references CumulativePercent (column C) data (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 3 — No LineChart sub-chart referencing column C (CumulativePercent) found")
        except Exception as e:
            print(f"ERROR: Component 3 — {e}")

        # Component 5: Chart title contains 'Pareto' (0.15 points)
        try:
            title_text = get_title_text(chart.title)
            if title_text and 'pareto' in title_text.lower():
                print(f"PASS: Component 5 — Chart title contains 'Pareto': '{title_text}' (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 5 — Chart title does not contain 'Pareto', found: '{title_text}'")
        except Exception as e:
            print(f"ERROR: Component 5 — {e}")

    # Component 4: Column D contains 80% reference line data (0.15 points)
    # Initial file has no data in column D; golden has D1='80% Reference' and D2:D11=0.8
    try:
        d_header = ws['D1'].value
        d_values = [ws.cell(row=r, column=4).value for r in range(2, 12)]
        # Check that D column has meaningful data (header + at least 8 of 10 values are 0.8)
        has_header = d_header is not None and str(d_header).strip() != ''
        count_08 = sum(1 for v in d_values if v is not None and abs(float(v) - 0.8) < 0.01)
        if has_header and count_08 >= 8:
            print(f"PASS: Component 4 — Column D has 80%% reference data: header='{d_header}', {count_08}/10 values are 0.8 (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — Column D missing or incomplete: header='{d_header}', {count_08}/10 values are 0.8")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
