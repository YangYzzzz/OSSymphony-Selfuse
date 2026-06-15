"""
Initial Setup: Create defect analysis spreadsheet for Pareto chart task
Task ID: calc_gcp_068
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_068'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'DefectAnalysis'

    # Headers
    headers = ['DefectType', 'Count', 'CumulativePercent']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows - defect types sorted descending by count
    data = [
        ['Missing Parts',    245, 0.252],
        ['Scratch',          189, 0.446],
        ['Dent',             156, 0.606],
        ['Color Mismatch',    98, 0.707],
        ['Misalignment',      87, 0.796],
        ['Electrical Fault',  62, 0.860],
        ['Packaging Damage',  45, 0.906],
        ['Label Error',       38, 0.945],
        ['Size Error',        28, 0.974],
        ['Other',             22, 1.000],
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        ws.cell(row=r, column=2, value=row_data[1])
        pct_cell = ws.cell(row=r, column=3, value=row_data[2])
        pct_cell.number_format = '0.0%'

    # Column widths for readability
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20

    # Light alternating row styling for data area
    light_fill = PatternFill(start_color='FFD6E4F0', end_color='FFD6E4F0', fill_type='solid')
    for r in range(2, 12):
        if r % 2 == 0:
            for c in range(1, 4):
                ws.cell(row=r, column=c).fill = light_fill

    # Thin borders on all data cells
    thin_side = Side(style='thin', color='000000')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    for r in range(1, 12):
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = thin_border

    # NO charts in initial state - the task is to create the Pareto chart

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
