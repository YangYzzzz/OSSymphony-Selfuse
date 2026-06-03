"""
Initial Setup: Add trusted macro file location task
Task ID: calc_mcp_021
Domain: libreoffice_calc

Creates a realistic spreadsheet and opens it in LibreOffice Calc.
No custom trusted file locations are configured (default state).
Also creates the /home/user/trusted_macros directory so it exists
as a valid path the user can add.
"""

import os
import shlex
import subprocess
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_021'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # --- Sheet 1: Project Tracker ---
    ws1 = wb.active
    ws1.title = 'Project Tracker'

    headers = ['Project Name', 'Lead', 'Department', 'Budget ($)', 'Start Date', 'Status']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    data = [
        ['Cloud Migration', 'Sarah Chen', 'Engineering', 245000, '2025-01-15', 'In Progress'],
        ['Brand Refresh', 'Marcus Johnson', 'Marketing', 78000, '2025-02-01', 'Planning'],
        ['ERP Upgrade', 'Priya Patel', 'Operations', 530000, '2024-11-10', 'In Progress'],
        ['Mobile App v3', 'David Kim', 'Engineering', 185000, '2025-03-01', 'Not Started'],
        ['Customer Portal', 'Lisa Wang', 'Product', 120000, '2025-01-20', 'In Progress'],
        ['Data Warehouse', 'James O\'Brien', 'Analytics', 310000, '2024-09-15', 'Testing'],
        ['Security Audit', 'Fatima Al-Rashid', 'IT Security', 95000, '2025-04-01', 'Planning'],
        ['Office Relocation', 'Tom Richards', 'Facilities', 450000, '2025-06-15', 'Not Started'],
        ['AI Chatbot', 'Nina Kowalski', 'Product', 165000, '2025-02-10', 'In Progress'],
        ['Payroll System', 'Carlos Mendez', 'HR', 88000, '2024-12-01', 'Completed'],
        ['Supply Chain Opt', 'Rachel Green', 'Operations', 275000, '2025-03-20', 'Planning'],
        ['Website Redesign', 'Alex Turner', 'Marketing', 62000, '2025-01-05', 'In Progress'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Set column widths
    ws1.column_dimensions['A'].width = 22
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 14
    ws1.column_dimensions['F'].width = 14

    # --- Sheet 2: Budget Summary ---
    ws2 = wb.create_sheet('Budget Summary')
    ws2_headers = ['Department', 'Total Budget', 'Spent', 'Remaining']
    for col, h in enumerate(ws2_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    budget_data = [
        ['Engineering', 430000, 215000, 215000],
        ['Marketing', 140000, 48000, 92000],
        ['Operations', 805000, 390000, 415000],
        ['Product', 285000, 142000, 143000],
        ['IT Security', 95000, 12000, 83000],
        ['HR', 88000, 88000, 0],
        ['Analytics', 310000, 280000, 30000],
        ['Facilities', 450000, 0, 450000],
    ]

    for r, row_data in enumerate(budget_data, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Create the trusted_macros directory so it exists as a valid path
    os.makedirs(f'{WORKDIR}/trusted_macros', exist_ok=True)
    print(f'Created directory: {WORKDIR}/trusted_macros')

    # GUI-ready startup: open LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
