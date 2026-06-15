"""
Reward Script: Add URL hyperlink in cell A3 of 'References' sheet with dark green font color
Task ID: calc_gg1_045
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): A3 display text is 'Calc Documentation'
  Component 2 (0.35): A3 has hyperlink to 'https://docs.libreoffice.org/calc'
  Component 3 (0.25): A3 font color is dark green (#006400)
  Component 4 (0.15): Existing hyperlinks in A1 and A2 are preserved
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_045'


def persist_app_state(domain: str):
    """Save any unsaved LibreOffice edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'References' sheet must exist
    if 'References' not in wb.sheetnames:
        print("FAIL: 'References' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['References']

    # Component 1: A3 display text is 'Calc Documentation' (0.25 points)
    try:
        a3_value = ws['A3'].value
        if a3_value is not None and str(a3_value).strip() == 'Calc Documentation':
            print(f"PASS: Component 1 — A3 display text is 'Calc Documentation' (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — expected 'Calc Documentation' in A3, found: {a3_value!r}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: A3 has hyperlink pointing to 'https://docs.libreoffice.org/calc' (0.35 points)
    try:
        hl = ws['A3'].hyperlink
        if hl is not None:
            target = hl.target
            if target is not None and target.rstrip('/') == 'https://docs.libreoffice.org/calc':
                print(f"PASS: Component 2 — A3 hyperlink target is '{target}' (0.35 pts)")
                total_score += 0.35
            else:
                print(f"FAIL: Component 2 — A3 hyperlink target is '{target}', expected 'https://docs.libreoffice.org/calc'")
        else:
            print("FAIL: Component 2 — A3 has no hyperlink")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: A3 font color is dark green (#006400 or close variant) (0.25 points)
    try:
        font_color = ws['A3'].font.color
        if font_color is not None and font_color.rgb is not None:
            rgb_str = str(font_color.rgb)
            # Accept both '00006400' (ARGB with alpha=00) and 'FF006400' (ARGB with alpha=FF)
            # The actual color portion is the last 6 chars
            color_hex = rgb_str[-6:].upper()
            if color_hex == '006400':
                print(f"PASS: Component 3 — A3 font color is dark green (rgb={rgb_str}) (0.25 pts)")
                total_score += 0.25
            else:
                # Also accept close variants of dark green
                # Parse RGB and check if it's reasonably close to #006400
                try:
                    r = int(color_hex[0:2], 16)
                    g = int(color_hex[2:4], 16)
                    b = int(color_hex[4:6], 16)
                    # Dark green: R close to 0, G close to 100 (0x64), B close to 0
                    if r <= 20 and 80 <= g <= 120 and b <= 20:
                        print(f"PASS: Component 3 — A3 font color is approximately dark green (#{color_hex}, rgb={rgb_str}) (0.25 pts)")
                        total_score += 0.25
                    else:
                        print(f"FAIL: Component 3 — A3 font color is #{color_hex} (rgb={rgb_str}), expected dark green (#006400)")
                except ValueError:
                    print(f"FAIL: Component 3 — could not parse font color: {rgb_str}")
        else:
            print("FAIL: Component 3 — A3 has no explicit font color set")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Existing hyperlinks in A1 and A2 are preserved (0.15 points)
    # This component only scores if A3 has ALSO been modified (i.e., at least one of
    # components 1-3 passed). Otherwise the initial file trivially has A1/A2 intact.
    try:
        if total_score > 0:
            a1_hl = ws['A1'].hyperlink
            a2_hl = ws['A2'].hyperlink
            a1_ok = a1_hl is not None and a1_hl.target is not None
            a2_ok = a2_hl is not None and a2_hl.target is not None
            if a1_ok and a2_ok:
                print(f"PASS: Component 4 — A1 hyperlink ({a1_hl.target}) and A2 hyperlink ({a2_hl.target}) preserved (0.15 pts)")
                total_score += 0.15
            else:
                missing = []
                if not a1_ok:
                    missing.append('A1')
                if not a2_ok:
                    missing.append('A2')
                print(f"FAIL: Component 4 — hyperlinks missing in: {', '.join(missing)}")
        else:
            # No task change detected, so A1/A2 preservation is a precondition, not a score
            print("SKIP: Component 4 — no task changes detected, skipping preservation check")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
