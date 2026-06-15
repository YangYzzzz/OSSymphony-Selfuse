"""
Initial Setup: Add a URL hyperlink in cell A3 with dark green font color
Task ID: calc_gg1_045
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_045'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: References ---
    ws1 = wb.active
    ws1.title = 'References'

    # Headers
    headers = ['Resource', 'Category', 'Added By', 'Date Added']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Row 2: Existing hyperlink - Getting Started Guide
    cell_a2 = ws1.cell(row=2, column=1, value='Getting Started Guide')
    cell_a2.hyperlink = 'https://wiki.documentfoundation.org/Documentation/GettingStarted'
    cell_a2.font = Font(name='Calibri', size=11, color='0000FF', underline='single')
    ws1.cell(row=2, column=2, value='Onboarding')
    ws1.cell(row=2, column=3, value='Rachel Torres')
    ws1.cell(row=2, column=4, value='2025-08-12')

    # Row 3: Existing hyperlink - Style Templates
    cell_a3 = ws1.cell(row=3, column=1, value='LibreOffice Templates')
    cell_a3.hyperlink = 'https://templates.libreoffice.org'
    cell_a3.font = Font(name='Calibri', size=11, color='0000FF', underline='single')
    ws1.cell(row=3, column=2, value='Templates')
    ws1.cell(row=3, column=3, value='David Kim')
    ws1.cell(row=3, column=4, value='2025-09-03')

    # Row 4 (A4): EMPTY — this is the reserved cell for the task
    # Note: task says A1 and A2 have hyperlinks, A3 is empty.
    # But openpyxl row numbering: row 1 = headers, so data rows start at 2.
    # Re-reading the task: "Cells A1 and A2 already have hyperlinks set up"
    # and "Cell A3 is empty". So A1 and A2 have hyperlinks, A3 is empty.
    # Let me fix: no header row concept for the hyperlink cells.
    # A1 = hyperlink, A2 = hyperlink, A3 = empty.

    # Let me redo: The task explicitly says A1 and A2 have hyperlinks, A3 is empty.
    # I'll put hyperlinks directly in A1 and A2, with supporting data in B/C/D columns.

    # Clear and rebuild
    for row in ws1.iter_rows(min_row=1, max_row=4, min_col=1, max_col=4):
        for cell in row:
            cell.value = None
            cell.font = Font()
            cell.fill = PatternFill()
            cell.alignment = Alignment()
            cell.hyperlink = None

    # Row 1: First hyperlink
    cell_a1 = ws1.cell(row=1, column=1, value='Getting Started Guide')
    cell_a1.hyperlink = 'https://wiki.documentfoundation.org/Documentation/GettingStarted'
    cell_a1.font = Font(name='Calibri', size=11, color='0000FF', underline='single')
    ws1.cell(row=1, column=2, value='Onboarding')
    ws1.cell(row=1, column=3, value='Rachel Torres')
    ws1.cell(row=1, column=4, value='2025-08-12')

    # Row 2: Second hyperlink
    cell_a2 = ws1.cell(row=2, column=1, value='LibreOffice Templates')
    cell_a2.hyperlink = 'https://templates.libreoffice.org'
    cell_a2.font = Font(name='Calibri', size=11, color='0000FF', underline='single')
    ws1.cell(row=2, column=2, value='Templates')
    ws1.cell(row=2, column=3, value='David Kim')
    ws1.cell(row=2, column=4, value='2025-09-03')

    # Row 3 (A3): EMPTY — reserved for the task
    # B3, C3, D3 can also be empty or have placeholder info
    ws1.cell(row=3, column=2, value='Internal Docs')
    ws1.cell(row=3, column=3, value='Pending')
    ws1.cell(row=3, column=4, value='')

    # Rows 4-8: Additional reference entries to add complexity
    more_refs = [
        ['Writer Handbook', 'https://wiki.documentfoundation.org/Documentation/WriterGuide',
         'Word Processing', 'Sarah Chen', '2025-10-01'],
        ['Impress Tips', 'https://wiki.documentfoundation.org/Documentation/ImpressGuide',
         'Presentations', 'Marcus Johnson', '2025-10-15'],
        ['Base Tutorial', 'https://wiki.documentfoundation.org/Documentation/BaseGuide',
         'Database', 'Elena Vasquez', '2025-11-02'],
        ['Draw Reference', 'https://wiki.documentfoundation.org/Documentation/DrawGuide',
         'Graphics', 'James Wright', '2025-11-20'],
        ['Macro Programming', 'https://wiki.documentfoundation.org/Documentation/MacroGuide',
         'Advanced', 'Anika Patel', '2025-12-05'],
    ]
    for i, (text, url, cat, author, date) in enumerate(more_refs, 4):
        cell = ws1.cell(row=i, column=1, value=text)
        cell.hyperlink = url
        cell.font = Font(name='Calibri', size=11, color='0000FF', underline='single')
        ws1.cell(row=i, column=2, value=cat)
        ws1.cell(row=i, column=3, value=author)
        ws1.cell(row=i, column=4, value=date)

    # Set column widths
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 14

    # --- Sheet 2: Team ---
    ws2 = wb.create_sheet('Team')
    team_headers = ['Name', 'Role', 'Department', 'Email', 'Start Date']
    for col, h in enumerate(team_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    team_data = [
        ['Rachel Torres', 'Documentation Lead', 'Engineering', 'r.torres@company.com', '2022-03-15'],
        ['David Kim', 'UX Designer', 'Product', 'd.kim@company.com', '2023-01-10'],
        ['Sarah Chen', 'Senior Developer', 'Engineering', 's.chen@company.com', '2021-07-22'],
        ['Marcus Johnson', 'Project Manager', 'Operations', 'm.johnson@company.com', '2022-09-01'],
        ['Elena Vasquez', 'Data Analyst', 'Analytics', 'e.vasquez@company.com', '2023-04-18'],
        ['James Wright', 'Graphic Designer', 'Creative', 'j.wright@company.com', '2023-06-30'],
        ['Anika Patel', 'Software Engineer', 'Engineering', 'a.patel@company.com', '2024-01-08'],
        ['Carlos Rivera', 'QA Engineer', 'Engineering', 'c.rivera@company.com', '2023-11-15'],
        ['Lisa Nakamura', 'HR Manager', 'People Ops', 'l.nakamura@company.com', '2021-02-01'],
        ['Tom Bradley', 'DevOps Engineer', 'Infrastructure', 't.bradley@company.com', '2022-12-05'],
    ]
    for r, row_data in enumerate(team_data, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 22
    ws2.column_dimensions['C'].width = 16
    ws2.column_dimensions['D'].width = 28
    ws2.column_dimensions['E'].width = 14

    # --- Sheet 3: Projects ---
    ws3 = wb.create_sheet('Projects')
    proj_headers = ['Project Name', 'Lead', 'Status', 'Budget ($)', 'Deadline']
    for col, h in enumerate(proj_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    proj_data = [
        ['Website Redesign', 'David Kim', 'In Progress', 45000, '2026-03-31'],
        ['API Migration', 'Sarah Chen', 'Planning', 82000, '2026-06-15'],
        ['Mobile App v2', 'Marcus Johnson', 'In Progress', 120000, '2026-05-01'],
        ['Data Pipeline', 'Elena Vasquez', 'Completed', 35000, '2025-12-31'],
        ['CI/CD Overhaul', 'Tom Bradley', 'In Progress', 28000, '2026-02-28'],
        ['Doc Portal', 'Rachel Torres', 'Planning', 15000, '2026-04-30'],
    ]
    for r, row_data in enumerate(proj_data, 2):
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r, column=c, value=val)

    ws3.column_dimensions['A'].width = 22
    ws3.column_dimensions['B'].width = 18
    ws3.column_dimensions['C'].width = 14
    ws3.column_dimensions['D'].width = 14
    ws3.column_dimensions['E'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
