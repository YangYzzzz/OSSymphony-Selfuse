"""
Initial Setup: Thesis bibliography spreadsheet with paper titles and years, missing DOIs/venues/URLs
Task ID: osworld_multi_apps_web_references_006
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user/Desktop'
TASK_ID = 'osworld_multi_apps_web_references_006'
FILENAME = 'thesis_refs.ods'
OUTPUT = f'{WORKDIR}/{FILENAME}'
XLSX_TMP = f'/home/user/{TASK_ID}_tmp.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    # Ensure Desktop directory exists
    os.makedirs(WORKDIR, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'References'

    # --- Header row ---
    headers = ['Title', 'Year', 'DOI', 'Venue', 'Open_Access_URL']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # --- Paper data: only Title and Year; DOI, Venue, Open_Access_URL are empty ---
    papers = [
        ('BERT: Pre-training of Deep Bidirectional Transformers for Language Understanding', 2019, '', '', ''),
        ('Language Models are Few-Shot Learners', 2020, '', '', ''),
        ('An Image is Worth 16x16 Words: Transformers for Image Recognition at Scale', 2021, '', '', ''),
        ('Denoising Diffusion Probabilistic Models', 2020, '', '', ''),
        ('Contrastive Multimodal Pretraining (CLIP)', 2021, '', '', ''),
        ('Masked Autoencoders Are Scalable Vision Learners', 2022, '', '', ''),
        ('PaLM: Scaling Language Modeling with Pathways', 2023, '', '', ''),
        ('Constitutional AI: Harmlessness from AI Feedback', 2022, '', '', ''),
    ]

    for row_idx, (title, year, doi, venue, url) in enumerate(papers, 2):
        ws.cell(row=row_idx, column=1, value=title)
        ws.cell(row=row_idx, column=2, value=year)
        # DOI, Venue, Open_Access_URL intentionally left empty (None / '')
        ws.cell(row=row_idx, column=3, value=None)
        ws.cell(row=row_idx, column=4, value=None)
        ws.cell(row=row_idx, column=5, value=None)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 70
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 50

    # Save as xlsx first, then convert to ods using LibreOffice CLI
    wb.save(XLSX_TMP)
    print(f'Temporary xlsx created: {XLSX_TMP}')

    # Convert xlsx to ods using LibreOffice headless
    convert_env = os.environ.copy()
    convert_env["DISPLAY"] = ":0"
    result = subprocess.run(
        ['libreoffice', '--headless', '--convert-to', 'ods', '--outdir', WORKDIR, XLSX_TMP],
        env=convert_env,
        capture_output=True,
        text=True,
        timeout=60
    )
    print(f'Conversion stdout: {result.stdout}')
    print(f'Conversion stderr: {result.stderr}')

    # The converted file will be named <task_id>_tmp.ods - rename to thesis_refs.ods
    converted = f'{WORKDIR}/{TASK_ID}_tmp.ods'
    if os.path.exists(converted):
        if os.path.exists(OUTPUT):
            os.remove(OUTPUT)
        os.rename(converted, OUTPUT)
        print(f'Renamed to: {OUTPUT}')
    elif not os.path.exists(OUTPUT):
        # Fallback: try saving directly as xlsx with ods extension via openpyxl
        # This won't be a proper ODS but LibreOffice can still open xlsx files
        wb.save(OUTPUT)
        print(f'Fallback: saved xlsx as {OUTPUT}')

    # Clean up temp file
    if os.path.exists(XLSX_TMP):
        os.remove(XLSX_TMP)

    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
