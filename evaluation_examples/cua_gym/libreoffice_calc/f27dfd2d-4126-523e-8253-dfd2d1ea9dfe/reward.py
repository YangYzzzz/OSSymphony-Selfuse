"""
Reward Script: Extract Innovation and Technology grant pass rates from PDF reports into LibreOffice Calc
Task ID: osworld_multi_apps_ecs_multi_report_012
Domain: libreoffice_calc
Scoring:
  - Component 1: Correct sheet name 'Innovation and Technology' exists (0.25 pts)
  - Component 2: Correct header row with 'University' and years 2020-2023 (0.25 pts)
  - Component 3: All 6 correct HK university names as row labels (0.25 pts)
  - Component 4: Correct pass rate values for all universities across all 4 years (0.25 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_ecs_multi_report_012'

# Expected values derived from task description and PDF source data
EXPECTED_SHEET_NAME = 'Innovation and Technology'

EXPECTED_UNIVERSITIES = [
    'University of Hong Kong',
    'Hong Kong Univ. of Sci. and Tech.',
    'Chinese Univ. of Hong Kong',
    'Hong Kong Polytechnic Univ.',
    'City University of Hong Kong',
    'Hong Kong Baptist University',
]

EXPECTED_YEARS = [2020, 2021, 2022, 2023]

# Expected pass rates as decimals (from PDF source data)
# {university_name: {year: pass_rate}}
EXPECTED_RATES = {
    'University of Hong Kong':          {2020: 0.60,  2021: 0.65,  2022: 0.70,  2023: 0.72},
    'Hong Kong Univ. of Sci. and Tech.':{2020: 0.75,  2021: 0.7826,2022: 0.80,  2023: 0.8214},
    'Chinese Univ. of Hong Kong':       {2020: 0.55,  2021: 0.5789,2022: 0.6190,2023: 0.65},
    'Hong Kong Polytechnic Univ.':      {2020: 0.50,  2021: 0.5238,2022: 0.55,  2023: 0.5833},
    'City University of Hong Kong':     {2020: 0.45,  2021: 0.4762,2022: 0.5238,2023: 0.55},
    'Hong Kong Baptist University':     {2020: 0.40,  2021: 0.4286,2022: 0.45,  2023: 0.48},
}


def normalize_rate(val):
    """Normalize a pass rate value to decimal form (0.0-1.0).
    Accepts both decimal (0.60) and percentage (60.0 or '60%') forms.
    """
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip().rstrip('%')
        try:
            val = float(val)
        except ValueError:
            return None
    try:
        val = float(val)
    except (ValueError, TypeError):
        return None
    # If value looks like a percentage (> 1.0), convert to decimal
    if val > 1.5:
        val = val / 100.0
    return val


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Gate: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Sheet named 'Innovation and Technology' exists (0.25 points)
    try:
        if EXPECTED_SHEET_NAME in wb.sheetnames:
            print(f"PASS: Component 1 — Sheet '{EXPECTED_SHEET_NAME}' found (0.25 pts)")
            total_score += 0.25
            ws = wb[EXPECTED_SHEET_NAME]
        else:
            # Try case-insensitive match
            matched = None
            for name in wb.sheetnames:
                if name.lower() == EXPECTED_SHEET_NAME.lower():
                    matched = name
                    break
            if matched:
                print(f"PASS (partial): Component 1 — Sheet found as '{matched}' (case mismatch) (0.25 pts)")
                total_score += 0.25
                ws = wb[matched]
            else:
                print(f"FAIL: Component 1 — Expected sheet '{EXPECTED_SHEET_NAME}', found sheets: {wb.sheetnames}")
                # Try to use the first sheet for further analysis
                ws = wb.worksheets[0] if wb.worksheets else None
                print(f"REWARD: {total_score}")
                return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 2: Header row contains 'University' and years 2020-2023 (0.25 points)
    try:
        header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        # Check year columns — accept both int and string versions
        header_years = []
        for val in header_row[1:]:
            if val is not None:
                try:
                    header_years.append(int(val))
                except (ValueError, TypeError):
                    pass

        has_university_header = (
            header_row[0] is not None and
            str(header_row[0]).strip().lower() in ('university', 'universities', 'university name')
        )
        has_year_headers = all(yr in header_years for yr in EXPECTED_YEARS)

        if has_university_header and has_year_headers:
            print(f"PASS: Component 2 — Header row correct: {header_row[:5]} (0.25 pts)")
            total_score += 0.25
        elif has_year_headers:
            print(f"PASS (partial): Component 2 — Year headers present (2020-2023), but university label may differ: '{header_row[0]}' (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — Expected headers with years 2020-2023, found: {header_row[:5]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: All 6 university names present as row labels (0.25 points)
    try:
        actual_universities = []
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val is not None and str(val).strip():
                actual_universities.append(str(val).strip())

        # Count how many expected universities are found (exact or close match)
        found_count = 0
        for expected in EXPECTED_UNIVERSITIES:
            # Exact match
            if expected in actual_universities:
                found_count += 1
            else:
                # Try case-insensitive match
                for actual in actual_universities:
                    if actual.lower() == expected.lower():
                        found_count += 1
                        break

        if found_count == len(EXPECTED_UNIVERSITIES):
            print(f"PASS: Component 3 — All {len(EXPECTED_UNIVERSITIES)} universities found as row labels (0.25 pts)")
            total_score += 0.25
        elif found_count >= len(EXPECTED_UNIVERSITIES) // 2:
            partial = 0.25 * (found_count / len(EXPECTED_UNIVERSITIES))
            print(f"PARTIAL: Component 3 — {found_count}/{len(EXPECTED_UNIVERSITIES)} universities found: partial {partial:.2f} pts")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {found_count}/{len(EXPECTED_UNIVERSITIES)} universities found")
            print(f"  Expected: {EXPECTED_UNIVERSITIES}")
            print(f"  Actual:   {actual_universities}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Correct pass rate values for all universities and all years (0.25 points)
    try:
        # Build a mapping of row index -> university name
        uni_row_map = {}
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val is not None and str(val).strip():
                uni_row_map[str(val).strip()] = row

        # Build a mapping of year -> column index (from header row)
        year_col_map = {}
        for col in range(2, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val is not None:
                try:
                    yr = int(val)
                    year_col_map[yr] = col
                except (ValueError, TypeError):
                    pass

        # Check each expected university x year combination
        total_checks = 0
        passed_checks = 0
        tolerance = 0.005  # Allow small floating-point differences

        for uni_name, year_rates in EXPECTED_RATES.items():
            # Find the actual row for this university (exact or case-insensitive)
            actual_row = uni_row_map.get(uni_name)
            if actual_row is None:
                for actual_uni, row in uni_row_map.items():
                    if actual_uni.lower() == uni_name.lower():
                        actual_row = row
                        break

            if actual_row is None:
                total_checks += len(year_rates)
                print(f"  MISS: University '{uni_name}' not found in spreadsheet")
                continue

            for year, expected_rate in year_rates.items():
                total_checks += 1
                actual_col = year_col_map.get(year)
                if actual_col is None:
                    print(f"  MISS: Year {year} column not found")
                    continue

                raw_val = ws.cell(row=actual_row, column=actual_col).value
                actual_rate = normalize_rate(raw_val)

                if actual_rate is not None and abs(actual_rate - expected_rate) <= tolerance:
                    passed_checks += 1
                else:
                    print(f"  FAIL: {uni_name} / {year}: expected ~{expected_rate:.4f}, got {raw_val!r} (normalized: {actual_rate})")

        if total_checks > 0:
            accuracy = passed_checks / total_checks
            component_score = 0.25 * accuracy
            if accuracy >= 0.95:
                print(f"PASS: Component 4 — {passed_checks}/{total_checks} pass rate values correct (0.25 pts)")
                total_score += 0.25
            elif accuracy >= 0.5:
                print(f"PARTIAL: Component 4 — {passed_checks}/{total_checks} values correct: {component_score:.3f} pts")
                total_score += component_score
            else:
                print(f"FAIL: Component 4 — Only {passed_checks}/{total_checks} values correct")
        else:
            print("FAIL: Component 4 — No data cells could be checked")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.3f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
