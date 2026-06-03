"""
Reward Script: Meeting Room Reservation Sheet Setup
Task ID: calc_ops_resource_room_reservation_038
Domain: libreoffice_calc
Scoring:
  Component 1: Data validation dropdown on RoomBookings!B2:B81 (0.25 pts)
  Component 2: Duration formula =(F-E)*24 in RoomBookings!G2:G81 (0.25 pts)
  Component 3: SUMIFS formula in RoomUtilization!C2:C16 (0.20 pts)
  Component 4: =C/8 formula with percentage format in RoomUtilization!D2:D16 (0.15 pts)
  Component 5: Conditional formatting on RoomUtilization!D2:D16 (3 rules: red/amber/green) (0.15 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_resource_room_reservation_038'


def normalize_formula(formula):
    """Normalize formula for comparison: strip leading '=', uppercase, remove spaces."""
    if not isinstance(formula, str):
        return ''
    f = formula.strip()
    if f.startswith('='):
        f = f[1:]
    return f.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion for meeting room reservation sheet setup.
    Returns a float between 0.0 and 1.0.

    The task required:
    1. Add data validation dropdown (Room 101, Room 102, Board Room) to RoomBookings!B2:B81
    2. Add duration formula =(F-E)*24 to RoomBookings!G2:G81
    3. Add SUMIFS formula to RoomUtilization!C2:C16 (sum booked hours per room per date)
    4. Add =C/8 formula with percentage format to RoomUtilization!D2:D16
    5. Add conditional formatting on D2:D16 (red >100%, amber 80-100%, green <80%)
    """
    total_score = 0.0

    # Load workbook — precondition gate
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify required sheets exist — precondition gate
    if 'RoomBookings' not in wb.sheetnames or 'RoomUtilization' not in wb.sheetnames:
        print("CRITICAL: Required sheets missing (RoomBookings, RoomUtilization)")
        print("REWARD: 0.0")
        return 0.0

    ws_rb = wb['RoomBookings']
    ws_ru = wb['RoomUtilization']

    # Component 1: Data validation dropdown on RoomBookings!B2:B81 (0.25 points)
    # Must have a list validation covering B2:B81 with options Room 101, Room 102, Board Room
    try:
        validations = ws_rb.data_validations.dataValidation
        found_dv = False
        dv_details = None
        for dv in validations:
            if dv.type == 'list':
                # Check the sqref covers B2:B81 (or a range containing it)
                sqref_str = str(dv.sqref)
                if 'B2' in sqref_str and 'B81' in sqref_str:
                    # Check the formula contains the three rooms
                    formula = dv.formula1 if dv.formula1 else ''
                    formula_clean = formula.replace('"', '').replace("'", '')
                    expected_rooms = ['Room 101', 'Room 102', 'Board Room']
                    rooms_found = all(room in formula_clean for room in expected_rooms)
                    if rooms_found:
                        found_dv = True
                        dv_details = formula
                        break

        if found_dv:
            print(f"PASS: Component 1 — Data validation dropdown found on B2:B81 with rooms: {dv_details} (0.25 pts)")
            total_score += 0.25
        else:
            # Try broader check: any list validation with correct rooms
            found_any = False
            for dv in validations:
                if dv.type == 'list':
                    formula = dv.formula1 if dv.formula1 else ''
                    formula_clean = formula.replace('"', '').replace("'", '')
                    expected_rooms = ['Room 101', 'Room 102', 'Board Room']
                    if all(room in formula_clean for room in expected_rooms):
                        found_any = True
                        dv_details = formula
                        print(f"PARTIAL: Component 1 — Dropdown found but sqref may not cover B2:B81; sqref={dv.sqref}, formula={formula}")
                        # Still award points — correct rooms, possibly slightly different range
                        total_score += 0.15
                        break
            if not found_any:
                print(f"FAIL: Component 1 — No list data validation with Room 101, Room 102, Board Room found on B2:B81")
                print(f"  Found validations: {[(dv.type, dv.formula1, str(dv.sqref)) for dv in validations]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Duration formula =(F-E)*24 in RoomBookings!G2:G81 (0.25 points)
    # All 80 rows must have a formula that computes (EndTime - StartTime) * 24
    try:
        formula_count = 0
        formula_correct_count = 0
        total_rows = 80  # rows 2 to 81

        for row in range(2, 82):
            cell_g = ws_rb.cell(row=row, column=7)
            val = cell_g.value
            if isinstance(val, str) and val.startswith('='):
                formula_count += 1
                # Check that the formula is =(F-E)*24 pattern (with row number)
                normalized = normalize_formula(val)
                expected = f'(F{row}-E{row})*24'
                if normalized == expected:
                    formula_correct_count += 1

        if formula_correct_count == total_rows:
            print(f"PASS: Component 2 — All {total_rows} duration formulas =(F-E)*24 present in G2:G81 (0.25 pts)")
            total_score += 0.25
        elif formula_correct_count >= total_rows * 0.9:
            print(f"PARTIAL: Component 2 — {formula_correct_count}/{total_rows} correct duration formulas found (0.15 pts)")
            total_score += 0.15
        elif formula_count >= total_rows * 0.9:
            print(f"PARTIAL: Component 2 — {formula_count} formulas present but pattern mismatch. Correct={formula_correct_count} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 — Expected 80 duration formulas in G2:G81, found {formula_count} formulas ({formula_correct_count} matching pattern)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: SUMIFS formula in RoomUtilization!C2:C16 (0.20 points)
    # All 15 rows must have a SUMIFS formula referencing RoomBookings!G, B, D columns
    try:
        sumifs_count = 0
        total_util_rows = 15  # rows 2 to 16

        for row in range(2, 17):
            cell_c = ws_ru.cell(row=row, column=3)
            val = cell_c.value
            if isinstance(val, str) and val.startswith('='):
                normalized = normalize_formula(val)
                # Check for SUMIFS referencing G and B column from RoomBookings
                if 'SUMIFS' in normalized and 'ROOMBOOKINGS' in normalized:
                    sumifs_count += 1

        if sumifs_count == total_util_rows:
            print(f"PASS: Component 3 — All {total_util_rows} SUMIFS formulas present in RoomUtilization!C2:C16 (0.20 pts)")
            total_score += 0.20
        elif sumifs_count >= total_util_rows * 0.9:
            print(f"PARTIAL: Component 3 — {sumifs_count}/{total_util_rows} SUMIFS formulas found (0.12 pts)")
            total_score += 0.12
        elif sumifs_count > 0:
            print(f"PARTIAL: Component 3 — Only {sumifs_count}/{total_util_rows} SUMIFS formulas found (0.06 pts)")
            total_score += 0.06
        else:
            print(f"FAIL: Component 3 — No SUMIFS formulas found in RoomUtilization!C2:C16")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: =C/8 formula with percentage format in RoomUtilization!D2:D16 (0.15 points)
    # All 15 rows must have =Cx/8 formula AND percentage number format
    try:
        formula_d_count = 0
        pct_format_count = 0
        total_util_rows = 15

        for row in range(2, 17):
            cell_d = ws_ru.cell(row=row, column=4)
            val = cell_d.value
            fmt = cell_d.number_format

            # Check formula pattern =Cx/8
            if isinstance(val, str) and val.startswith('='):
                normalized = normalize_formula(val)
                expected = f'C{row}/8'
                if normalized == expected:
                    formula_d_count += 1

            # Check percentage format (various percentage formats are acceptable)
            if fmt and ('%' in fmt):
                pct_format_count += 1

        if formula_d_count == total_util_rows and pct_format_count == total_util_rows:
            print(f"PASS: Component 4 — All {total_util_rows} =C/8 formulas with percentage format in D2:D16 (0.15 pts)")
            total_score += 0.15
        elif formula_d_count == total_util_rows:
            print(f"PARTIAL: Component 4 — {formula_d_count} =C/8 formulas present but only {pct_format_count} have percentage format (0.10 pts)")
            total_score += 0.10
        elif formula_d_count >= total_util_rows * 0.9 and pct_format_count >= total_util_rows * 0.9:
            print(f"PARTIAL: Component 4 — {formula_d_count}/{total_util_rows} formulas, {pct_format_count}/{total_util_rows} pct format (0.10 pts)")
            total_score += 0.10
        elif formula_d_count > 0:
            print(f"PARTIAL: Component 4 — Only {formula_d_count}/{total_util_rows} =C/8 formulas found (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4 — No =C/8 formulas with percentage format in RoomUtilization!D2:D16. formula_count={formula_d_count}, pct_count={pct_format_count}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Conditional formatting on RoomUtilization!D2:D16 (0.15 points)
    # Must have 3 rules: red if >100% (>1), amber if 80-100% (0.8-1), green if <80% (<0.8)
    try:
        cf_rules = ws_ru.conditional_formatting
        found_red = False    # >1 with red fill
        found_amber = False  # between 0.8 and 1 with amber/yellow fill
        found_green = False  # <0.8 with green fill

        for cf in cf_rules:
            for rule in cf.rules:
                try:
                    rtype = rule.type
                    rformula = getattr(rule, 'formula', None)
                    roperator = None
                    try:
                        roperator = rule.operator
                    except Exception:
                        pass

                    # Check fill color
                    fill_rgb = None
                    try:
                        if rule.dxf and rule.dxf.fill and rule.dxf.fill.fgColor:
                            fill_rgb = rule.dxf.fill.fgColor.rgb
                    except Exception:
                        pass

                    if rtype == 'cellIs' and roperator == 'greaterThan':
                        # Check for >1 rule (>100%)
                        if rformula and '1' in str(rformula):
                            # Red fill: FFFF0000 or similar red
                            if fill_rgb and ('FF0000' in fill_rgb or fill_rgb == 'FFFF0000'):
                                found_red = True
                                print(f"  CF: Found red rule (>1): operator={roperator}, formula={rformula}, fill={fill_rgb}")

                    if rtype == 'cellIs' and roperator == 'between':
                        # Check for 0.8-1 range (80%-100%)
                        if rformula and len(rformula) >= 2:
                            # Amber/yellow fill: FFFFC000 or similar
                            if fill_rgb and ('FFC000' in fill_rgb or 'FFFF00' in fill_rgb or fill_rgb == 'FFFFC000'):
                                found_amber = True
                                print(f"  CF: Found amber rule (between): operator={roperator}, formula={rformula}, fill={fill_rgb}")

                    if rtype == 'cellIs' and roperator == 'lessThan':
                        # Check for <0.8 rule (<80%)
                        if rformula and '0.8' in str(rformula):
                            # Green fill: FF00B050 or similar green
                            if fill_rgb and ('00B050' in fill_rgb or '00FF00' in fill_rgb or '008000' in fill_rgb or fill_rgb == 'FF00B050'):
                                found_green = True
                                print(f"  CF: Found green rule (<0.8): operator={roperator}, formula={rformula}, fill={fill_rgb}")
                except Exception as e2:
                    print(f"  CF rule parse error: {e2}")

        rules_found = sum([found_red, found_amber, found_green])

        if found_red and found_amber and found_green:
            print(f"PASS: Component 5 — All 3 conditional formatting rules present (red/amber/green on D2:D16) (0.15 pts)")
            total_score += 0.15
        elif rules_found == 2:
            print(f"PARTIAL: Component 5 — {rules_found}/3 CF rules found (red={found_red}, amber={found_amber}, green={found_green}) (0.08 pts)")
            total_score += 0.08
        elif rules_found == 1:
            print(f"PARTIAL: Component 5 — {rules_found}/3 CF rules found (0.04 pts)")
            total_score += 0.04
        else:
            print(f"FAIL: Component 5 — No conditional formatting rules found on D2:D16 (red={found_red}, amber={found_amber}, green={found_green})")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
