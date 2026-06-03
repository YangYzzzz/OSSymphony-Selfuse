"""
Initial Setup: Meeting Room Reservation Sheet
Task ID: calc_ops_resource_room_reservation_038
Domain: libreoffice_calc
"""

import openpyxl
from datetime import date, time, timedelta
import datetime

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_resource_room_reservation_038'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: RoomBookings ---
    ws1 = wb.active
    ws1.title = 'RoomBookings'

    # Headers for RoomBookings
    headers_rb = ['Booking ID', 'Room', 'Organizer', 'Date', 'Start Time', 'End Time',
                  'Duration Hours', 'Attendees', 'Purpose']
    for col, h in enumerate(headers_rb, 1):
        ws1.cell(row=1, column=col, value=h)

    # Rooms and organizers for realistic data
    rooms = ['Room 101', 'Room 102', 'Board Room']
    organizers = [
        'Sarah Chen', 'Marcus Johnson', 'Emily Rodriguez', 'David Kim',
        'Priya Patel', 'James Wright', 'Aisha Okonkwo', 'Liam Fitzgerald',
        'Mei-Lin Zhao', 'Carlos Reyes', 'Anna Bergstrom', 'Tobias Muller',
        'Fatima Al-Hassan', 'Ravi Sharma', 'Sophie Dubois'
    ]
    purposes = [
        'Weekly Standup', 'Sprint Planning', 'Client Presentation', 'Budget Review',
        'Product Roadmap Discussion', 'HR Interview', 'Team Retrospective',
        'Vendor Meeting', 'Training Session', 'Quarterly Review',
        'Project Kickoff', 'Board Meeting', 'Technical Design Review',
        'Sales Demo', 'Strategy Workshop', 'All Hands Meeting',
        'Performance Review', 'Design Review', 'Operations Sync'
    ]

    # Generate 80 bookings with realistic dates, times
    import random
    random.seed(42)  # Reproducible
    base_date = date(2025, 3, 3)  # Monday

    bookings = []
    for i in range(80):
        booking_id = f'BK{1001 + i}'
        room = rooms[i % 3]
        organizer = organizers[i % len(organizers)]
        # Spread across ~20 working days (4 weeks)
        day_offset = (i // 4) % 20
        booking_date = base_date + timedelta(days=day_offset + (day_offset // 5) * 2)  # skip weekends
        # Vary start times between 8am and 5pm
        start_hour = 8 + (i % 9)
        start_min = [0, 30][i % 2]
        duration_slots = [1, 1.5, 2, 0.5][i % 4]
        end_hour = start_hour + int(duration_slots)
        end_min = start_min + int((duration_slots % 1) * 60)
        if end_min >= 60:
            end_hour += 1
            end_min -= 60

        start_t = time(start_hour, start_min)
        end_t = time(min(end_hour, 17), end_min)

        attendees = 3 + (i % 12)
        purpose = purposes[i % len(purposes)]

        bookings.append([booking_id, room, organizer, booking_date, start_t, end_t,
                         None, attendees, purpose])

    for r, row_data in enumerate(bookings, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)
            # Format date column
            if c == 4 and val is not None:
                ws1.cell(row=r, column=c).number_format = 'yyyy-mm-dd'
            # Format time columns
            if c in (5, 6) and val is not None:
                ws1.cell(row=r, column=c).number_format = 'hh:mm'

    # --- Sheet 2: RoomUtilization ---
    ws2 = wb.create_sheet('RoomUtilization')

    headers_ru = ['Room', 'Date', 'Total Booked Hours', 'Utilization %']
    for col, h in enumerate(headers_ru, 1):
        ws2.cell(row=1, column=col, value=h)

    # 15 room-date combinations: 5 dates x 3 rooms
    util_dates = [
        date(2025, 3, 3), date(2025, 3, 4), date(2025, 3, 5),
        date(2025, 3, 6), date(2025, 3, 7)
    ]
    util_rooms = ['Room 101', 'Room 102', 'Board Room']
    util_rows = []
    for d in util_dates:
        for rm in util_rooms:
            util_rows.append([rm, d, None, None])

    for r, row_data in enumerate(util_rows, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)
            if c == 2 and val is not None:
                ws2.cell(row=r, column=c).number_format = 'yyyy-mm-dd'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'RoomBookings rows: {ws1.max_row - 1}')
    print(f'RoomUtilization rows: {ws2.max_row - 1}')


create_initial()
