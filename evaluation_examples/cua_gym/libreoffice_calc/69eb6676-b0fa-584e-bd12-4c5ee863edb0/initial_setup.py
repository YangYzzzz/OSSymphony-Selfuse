"""
Initial Setup: Event Calendar with empty date column needing validation
Task ID: calc_gcv_059
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_059'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Event_Calendar"

    # --- Headers ---
    headers = ["Event ID", "Event Name", "Venue", "Organizer", "Event Date"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="000000")
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # --- 24 Event rows (rows 2-25) ---
    events = [
        ["EVT-001", "Annual Tech Summit", "Grand Hyatt Convention Center", "Sarah Chen"],
        ["EVT-002", "Marketing Strategy Workshop", "Hilton Downtown Ballroom", "Marcus Johnson"],
        ["EVT-003", "Employee Wellness Fair", "Corporate Campus Atrium", "Priya Sharma"],
        ["EVT-004", "Q1 Sales Review Meeting", "Conference Room 4A", "David Kim"],
        ["EVT-005", "Product Launch Gala", "Ritz-Carlton Grand Hall", "Olivia Martinez"],
        ["EVT-006", "Board of Directors Retreat", "Lakeside Resort", "James O'Brien"],
        ["EVT-007", "Customer Appreciation Dinner", "The Capital Grille", "Aisha Patel"],
        ["EVT-008", "New Hire Orientation", "Training Center B2", "Robert Williams"],
        ["EVT-009", "Innovation Hackathon", "Tech Hub Workspace", "Mei-Lin Wang"],
        ["EVT-010", "Quarterly Town Hall", "Main Auditorium", "Thomas Anderson"],
        ["EVT-011", "Charity Golf Tournament", "Pine Valley Golf Club", "Jennifer Brooks"],
        ["EVT-012", "Supply Chain Summit", "Marriott Conference Wing", "Ahmed Hassan"],
        ["EVT-013", "Leadership Development Seminar", "Executive Boardroom", "Catherine Lee"],
        ["EVT-014", "IT Security Awareness Day", "Virtual + Room 301", "Raj Gupta"],
        ["EVT-015", "Partner Network Mixer", "Sky Lounge Rooftop Bar", "Nicole Foster"],
        ["EVT-016", "Budget Planning Workshop", "Finance Meeting Room", "William Chang"],
        ["EVT-017", "Annual Holiday Party", "Riverside Event Center", "Samantha Davis"],
        ["EVT-018", "Cross-Department Sync", "Conference Room 2C", "Michael Torres"],
        ["EVT-019", "Vendor Showcase Exhibition", "Exhibition Hall East", "Laura Nguyen"],
        ["EVT-020", "Diversity & Inclusion Forum", "Community Center", "Derek Robinson"],
        ["EVT-021", "Engineering Sprint Review", "Dev Lab Alpha", "Yuki Tanaka"],
        ["EVT-022", "Client Onboarding Session", "Meeting Room 5B", "Patricia Walsh"],
        ["EVT-023", "Sustainability Initiative Launch", "Green Conference Hall", "Carlos Mendoza"],
        ["EVT-024", "Year-End Awards Ceremony", "Grand Ballroom", "Elizabeth Park"],
    ]

    data_font = Font(name="Calibri", size=11)
    data_align = Alignment(vertical="center")
    date_format = 'yyyy-mm-dd'

    for r, row_data in enumerate(events, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = header_border

        # Column E (Event Date) - empty but formatted as date
        date_cell = ws.cell(row=r, column=5)
        date_cell.number_format = date_format
        date_cell.font = data_font
        date_cell.alignment = data_align
        date_cell.border = header_border

    # --- Column widths ---
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 16

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f"Initial file created: {OUTPUT}")

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print("GUI_READY: launched LibreOffice Calc with DISPLAY=:0")


create_initial()
