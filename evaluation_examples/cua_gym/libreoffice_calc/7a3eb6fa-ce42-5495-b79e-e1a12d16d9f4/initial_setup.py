"""
Initial Setup: Expense Report spreadsheet with no formatting applied
Task ID: calc_fmt_comprehensive_report_styling_100
Domain: libreoffice_calc

Creates an expense report with:
- One sheet named 'Expense Report'
- Row 1: headers (no formatting)
- Rows 2-20: 19 expense records
- Column B: date serial values (no date format applied)
- Columns C, D, E: numeric dollar amounts (no currency format applied)
- No borders, no fills, no bold, no special formatting
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_comprehensive_report_styling_100'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Expense Report'

    # Row 1: Headers (plain text, no formatting)
    headers = ['Expense ID', 'Date', 'Category Amount', 'Receipts Total', 'Approved Amount', 'Status']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Rows 2-20: Realistic expense records
    # Column B: Excel date serial values (days since 1900-01-01)
    # 2024-01-05 = 45296, 2024-01-10 = 45301, etc.
    data = [
        # ExpID, DateSerial,   CatAmt,    RcptTotal, ApprAmt,   Status
        ['EXP-001', 45296,   1250.00,   1250.00,   1250.00,   'Approved'],
        ['EXP-002', 45310,   5800.50,   5700.00,   5700.00,   'Approved'],
        ['EXP-003', 45325,    320.75,    320.75,    320.75,   'Approved'],
        ['EXP-004', 45338,   6450.00,   6200.00,   6200.00,   'Pending'],
        ['EXP-005', 45352,   2100.00,   2100.00,   2100.00,   'Approved'],
        ['EXP-006', 45366,    890.25,    890.25,    890.25,   'Approved'],
        ['EXP-007', 45380,   7350.00,   7000.00,   7000.00,   'Approved'],
        ['EXP-008', 45394,   3200.00,   3150.00,   3150.00,   'Pending'],
        ['EXP-009', 45408,   4500.00,   4500.00,   4500.00,   'Approved'],
        ['EXP-010', 45422,   1850.75,   1850.75,   1850.75,   'Approved'],
        ['EXP-011', 45436,   9200.00,   9000.00,   8500.00,   'Under Review'],
        ['EXP-012', 45450,    675.50,    675.50,    675.50,   'Approved'],
        ['EXP-013', 45464,   5100.00,   4950.00,   4950.00,   'Approved'],
        ['EXP-014', 45478,   2780.00,   2780.00,   2780.00,   'Approved'],
        ['EXP-015', 45492,   6800.00,   6500.00,   6500.00,   'Pending'],
        ['EXP-016', 45506,   1420.25,   1420.25,   1420.25,   'Approved'],
        ['EXP-017', 45520,   3950.00,   3800.00,   3800.00,   'Approved'],
        ['EXP-018', 45534,   8100.00,   7800.00,   7800.00,   'Under Review'],
        ['EXP-019', 45548,    540.00,    540.00,    540.00,   'Approved'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # No formatting applied - all cells use default 'General' format
    # No borders, no fills, no bold, no colors

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Expense Report')
    print(f'  Rows: 1 header + 19 data rows')
    print(f'  Columns: Expense ID, Date (serial), Category Amount, Receipts Total, Approved Amount, Status')
    print(f'  Formatting: NONE (all General format)')


create_initial()
