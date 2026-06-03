"""
Reward Script: Apply comprehensive professional styling to expense report
Task ID: calc_fmt_comprehensive_report_styling_100
Domain: libreoffice_calc
Scoring:
  - Component 1: Header row A1:F1 bold 14pt, dark blue fill (#002060), white text     (0.30 pts)
  - Component 2: Currency format '$#,##0.00' for C2:E20                               (0.20 pts)
  - Component 3: Date format 'DD-MMM-YYYY' for B2:B20                                 (0.15 pts)
  - Component 4: Thick outer border around A1:F20                                     (0.20 pts)
  - Component 5: Thin inner borders for all cells in A1:F20                           (0.05 pts)
  - Component 6: Conditional formatting on E2:E20 (value > 5000 -> fill #FFC7CE)     (0.10 pts)
  Total: 1.00
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmt_comprehensive_report_styling_100'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet exists
    if 'Expense Report' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Expense Report' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Expense Report']

    # -------------------------------------------------------------------------
    # Component 1: Header row A1:F1 — bold=True, size=14, fill=#002060, font color=#FFFFFF (0.30 pts)
    # All 6 header cells must have: bold=True, size=14, fgColor=FF002060, fontColor=FFFFFFFF
    # This FAILS on initial (no formatting) → PASSES on golden
    # -------------------------------------------------------------------------
    try:
        header_pass_count = 0
        header_cols = ['A', 'B', 'C', 'D', 'E', 'F']
        for col in header_cols:
            cell = ws[f'{col}1']
            bold_ok = cell.font.bold is True
            size_ok = cell.font.size is not None and abs(float(cell.font.size) - 14.0) < 0.5
            try:
                fill_ok = cell.fill.fgColor.rgb == 'FF002060'
            except Exception:
                fill_ok = False
            try:
                # Font color white: FFFFFFFF
                font_color_ok = cell.font.color.rgb == 'FFFFFFFF'
            except Exception:
                font_color_ok = False
            if bold_ok and size_ok and fill_ok and font_color_ok:
                header_pass_count += 1

        if header_pass_count == 6:
            print(f"PASS: Component 1 — All 6 header cells styled correctly (bold 14pt, #002060 fill, white text) (0.30 pts)")
            total_score += 0.30
        elif header_pass_count >= 3:
            print(f"PARTIAL: Component 1 — {header_pass_count}/6 header cells styled correctly (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — Only {header_pass_count}/6 header cells styled correctly. Expected bold=True, size=14, fill=FF002060, font_color=FFFFFFFF")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Currency format '$#,##0.00' for columns C, D, E (rows 2-20) (0.20 pts)
    # This FAILS on initial (format=General) → PASSES on golden
    # -------------------------------------------------------------------------
    try:
        currency_pass_count = 0
        total_currency_cells = 0
        for row in range(2, 21):
            for col in ['C', 'D', 'E']:
                cell = ws[f'{col}{row}']
                total_currency_cells += 1
                if cell.number_format == '$#,##0.00':
                    currency_pass_count += 1

        if currency_pass_count == total_currency_cells:
            print(f"PASS: Component 2 — All {total_currency_cells} cells in C2:E20 have currency format '$#,##0.00' (0.20 pts)")
            total_score += 0.20
        elif currency_pass_count >= total_currency_cells // 2:
            print(f"PARTIAL: Component 2 — {currency_pass_count}/{total_currency_cells} currency cells correct (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 — Only {currency_pass_count}/{total_currency_cells} cells in C2:E20 have currency format. Expected '$#,##0.00'")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Date format 'DD-MMM-YYYY' for column B (rows 2-20) (0.15 pts)
    # This FAILS on initial (format=General) → PASSES on golden
    # -------------------------------------------------------------------------
    try:
        date_pass_count = 0
        total_date_cells = 19  # B2:B20
        for row in range(2, 21):
            cell = ws[f'B{row}']
            if cell.number_format == 'DD-MMM-YYYY':
                date_pass_count += 1

        if date_pass_count == total_date_cells:
            print(f"PASS: Component 3 — All {total_date_cells} cells in B2:B20 have date format 'DD-MMM-YYYY' (0.15 pts)")
            total_score += 0.15
        elif date_pass_count >= total_date_cells // 2:
            print(f"PARTIAL: Component 3 — {date_pass_count}/{total_date_cells} date cells correct (0.07 pts)")
            total_score += 0.07
        else:
            print(f"FAIL: Component 3 — Only {date_pass_count}/{total_date_cells} cells in B2:B20 have date format. Expected 'DD-MMM-YYYY', found {ws['B2'].number_format}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: Thick outer border around A1:F20 (0.20 pts)
    # Outer edges: top row A1:F1 (top=thick), bottom row A20:F20 (bottom=thick),
    # left col A1:A20 (left=thick), right col F1:F20 (right=thick)
    # This FAILS on initial (no borders) → PASSES on golden
    # -------------------------------------------------------------------------
    try:
        thick_border_issues = []

        # Top edge: row 1 must have top=thick
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            b = ws[f'{col}1'].border
            if b.top.style != 'thick':
                thick_border_issues.append(f"{col}1 top={b.top.style} (expected thick)")

        # Bottom edge: row 20 must have bottom=thick
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            b = ws[f'{col}20'].border
            if b.bottom.style != 'thick':
                thick_border_issues.append(f"{col}20 bottom={b.bottom.style} (expected thick)")

        # Left edge: col A must have left=thick
        for row in range(1, 21):
            b = ws[f'A{row}'].border
            if b.left.style != 'thick':
                thick_border_issues.append(f"A{row} left={b.left.style} (expected thick)")

        # Right edge: col F must have right=thick
        for row in range(1, 21):
            b = ws[f'F{row}'].border
            if b.right.style != 'thick':
                thick_border_issues.append(f"F{row} right={b.right.style} (expected thick)")

        if not thick_border_issues:
            print(f"PASS: Component 4 — Thick outer border correctly applied to entire A1:F20 range (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 4 — Thick outer border issues ({len(thick_border_issues)} problems). Examples: {thick_border_issues[:3]}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -------------------------------------------------------------------------
    # Component 5: Thin inner borders for all cells in A1:F20 (0.05 pts)
    # All cells must have at least thin borders on all sides (inner cells)
    # We check a sample of interior cells not on the outer edge
    # This FAILS on initial (no borders) → PASSES on golden
    # -------------------------------------------------------------------------
    try:
        inner_border_issues = []
        # Check representative interior cells (not on outer edges)
        interior_cells = ['B5', 'C10', 'D15', 'E3', 'B19', 'C7', 'D12']
        for coord in interior_cells:
            b = ws[coord].border
            # All 4 sides must have at least 'thin'
            for side_name, side_val in [('left', b.left), ('right', b.right), ('top', b.top), ('bottom', b.bottom)]:
                if side_val.style not in ('thin', 'medium', 'thick'):
                    inner_border_issues.append(f"{coord} {side_name}={side_val.style}")

        if not inner_border_issues:
            print(f"PASS: Component 5 — Thin inner borders correctly applied to interior cells (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — Inner border issues: {inner_border_issues[:3]}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # -------------------------------------------------------------------------
    # Component 6: Conditional formatting on E2:E20 — value > 5000 -> fill #FFC7CE (0.10 pts)
    # This FAILS on initial (no CF) → PASSES on golden
    # -------------------------------------------------------------------------
    try:
        cf_ranges_on_e = []
        for cf_range in ws.conditional_formatting:
            range_str = str(cf_range)
            if 'E2:E20' in range_str or 'E2' in range_str:
                cf_ranges_on_e.append(cf_range)

        if not cf_ranges_on_e:
            print(f"FAIL: Component 6 — No conditional formatting found on E2:E20")
        else:
            # Evaluate each matching CF range for correct configuration
            cf_details = []
            for cf_range in cf_ranges_on_e:
                for rule in cf_range.rules:
                    if rule.type == 'cellIs' and rule.operator == 'greaterThan':
                        formula_val = rule.formula[0] if rule.formula else None
                        try:
                            formula_num = float(formula_val)
                            formula_matches = abs(formula_num - 5000.0) < 0.01
                        except (TypeError, ValueError):
                            formula_matches = False
                        try:
                            fill_color = rule.dxf.fill.fgColor.rgb
                            color_matches = (fill_color == 'FFFFC7CE')
                        except Exception:
                            color_matches = False
                            fill_color = 'unknown'
                        cf_details.append({
                            'formula_matches': formula_matches,
                            'color_matches': color_matches,
                            'formula_val': formula_val,
                            'fill_color': fill_color if not color_matches else 'FFFFC7CE'
                        })

            # Check if any rule matches all criteria
            all_criteria_met = any(
                d['formula_matches'] and d['color_matches']
                for d in cf_details
            )
            if all_criteria_met:
                print(f"PASS: Component 6 — Conditional formatting on E2:E20 correctly configured (>5000 fill #FFC7CE) (0.10 pts)")
                total_score += 0.10
            elif cf_details:
                first = cf_details[0]
                print(f"FAIL: Component 6 — CF rule found on E column but details incorrect. "
                      f"formula_matches={first['formula_matches']} (val={first['formula_val']}), "
                      f"color_matches={first['color_matches']} (color={first.get('fill_color','?')})")
            else:
                print(f"FAIL: Component 6 — CF range found on E column but no 'cellIs greaterThan' rule detected")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
