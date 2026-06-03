"""
Reward Script: Onboarding checklist tracker with percentage completion and days-until-deadline calculations.
Task ID: calc_hr_062
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): D2:D6 Status formulas — =IF(C="Yes","Done",IF(B<TODAY(),"Overdue","Pending"))
  Component 2 (0.3): E2:E6 Days Remaining formulas — =IF(C="Yes","",B-TODAY())
  Component 3 (0.3): G2 Completion % formula — =COUNTIF(C2:C6,"Yes")/COUNTA(C2:C6)
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_062'


def normalize_formula(f):
    """Normalize formula for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def check_status_formula(formula_str, row):
    """
    Check if a D-column formula matches the expected pattern:
    =IF(C<row>="Yes","Done",IF(B<row><TODAY(),"Overdue","Pending"))
    Allow minor variants (quotes, spacing).
    """
    norm = normalize_formula(formula_str)
    # Expected pattern (normalized): =IF(C<row>="YES","DONE",IF(B<row><TODAY(),"OVERDUE","PENDING"))
    expected = f'=IF(C{row}="YES","DONE",IF(B{row}<TODAY(),"OVERDUE","PENDING"))'
    return norm == normalize_formula(expected)


def check_days_remaining_formula(formula_str, row):
    """
    Check if an E-column formula matches the expected pattern:
    =IF(C<row>="Yes","",B<row>-TODAY())
    """
    norm = normalize_formula(formula_str)
    expected = f'=IF(C{row}="YES","",B{row}-TODAY())'
    return norm == normalize_formula(expected)


def check_completion_formula(formula_str):
    """
    Check if G2 formula matches: =COUNTIF(C2:C6,"Yes")/COUNTA(C2:C6)
    """
    norm = normalize_formula(formula_str)
    expected = '=COUNTIF(C2:C6,"YES")/COUNTA(C2:C6)'
    return norm == normalize_formula(expected)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the Onboarding sheet exists
    if 'Onboarding' not in wb.sheetnames:
        print("FAIL: 'Onboarding' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Onboarding']

    # Component 1: D2:D6 Status formulas (0.4 points)
    # Each correct formula earns 0.08 points (5 rows x 0.08 = 0.4)
    try:
        status_score = 0.0
        status_details = []
        for row in range(2, 7):
            cell_val = ws.cell(row=row, column=4).value  # D column
            if cell_val and isinstance(cell_val, str) and cell_val.startswith('='):
                if check_status_formula(cell_val, row):
                    status_score += 0.08
                    status_details.append(f"D{row}: PASS")
                else:
                    status_details.append(f"D{row}: WRONG formula '{cell_val}'")
            else:
                status_details.append(f"D{row}: NO formula (found: {repr(cell_val)})")

        if status_score > 0:
            print(f"PASS: Component 1 — Status formulas ({status_score:.2f} pts) [{', '.join(status_details)}]")
            total_score += status_score
        else:
            print(f"FAIL: Component 1 — No valid Status formulas [{', '.join(status_details)}]")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: E2:E6 Days Remaining formulas (0.3 points)
    # Each correct formula earns 0.06 points (5 rows x 0.06 = 0.3)
    try:
        days_score = 0.0
        days_details = []
        for row in range(2, 7):
            cell_val = ws.cell(row=row, column=5).value  # E column
            if cell_val and isinstance(cell_val, str) and cell_val.startswith('='):
                if check_days_remaining_formula(cell_val, row):
                    days_score += 0.06
                    days_details.append(f"E{row}: PASS")
                else:
                    days_details.append(f"E{row}: WRONG formula '{cell_val}'")
            else:
                days_details.append(f"E{row}: NO formula (found: {repr(cell_val)})")

        if days_score > 0:
            print(f"PASS: Component 2 — Days Remaining formulas ({days_score:.2f} pts) [{', '.join(days_details)}]")
            total_score += days_score
        else:
            print(f"FAIL: Component 2 — No valid Days Remaining formulas [{', '.join(days_details)}]")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: G2 Completion % formula (0.3 points)
    try:
        g2_val = ws['G2'].value
        if g2_val and isinstance(g2_val, str) and g2_val.startswith('='):
            if check_completion_formula(g2_val):
                print(f"PASS: Component 3 — G2 Completion % formula correct ({g2_val}) (0.30 pts)")
                total_score += 0.30
            else:
                print(f"FAIL: Component 3 — G2 formula mismatch: {g2_val}")
        else:
            print(f"FAIL: Component 3 — G2 has no formula (found: {repr(g2_val)})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
