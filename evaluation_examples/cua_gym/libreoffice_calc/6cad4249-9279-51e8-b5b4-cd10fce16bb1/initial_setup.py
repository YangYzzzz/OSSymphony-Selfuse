"""
Initial Setup: Build an automated onboarding checklist tracker
Task ID: calc_hr_062
Domain: libreoffice_calc

Creates a spreadsheet with onboarding task data but WITHOUT the formulas
for Status (D), Days Remaining (E), and Completion % (G2).
The agent must add those formulas.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_062'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Onboarding'

    # --- Headers ---
    headers = {
        'A1': 'Task',
        'B1': 'Due Date',
        'C1': 'Completed',
        'D1': 'Status',
        'E1': 'Days Remaining',
        'G1': 'Completion %',
    }
    header_font = Font(bold=True)
    for coord, label in headers.items():
        cell = ws[coord]
        cell.value = label
        cell.font = header_font

    # --- Data rows ---
    tasks = [
        ('IT Setup',             '2024-03-05', 'Yes'),
        ('Badge Photo',          '2024-03-07', 'No'),
        ('Benefits Enrollment',  '2024-03-15', 'No'),
        ('Safety Training',      '2024-03-10', 'Yes'),
        ('Intro Meetings',       '2024-03-12', 'No'),
    ]

    for i, (task_name, due_str, completed) in enumerate(tasks, start=2):
        ws.cell(row=i, column=1, value=task_name)
        # Store due dates as actual date objects for proper formula use
        y, m, d = due_str.split('-')
        ws.cell(row=i, column=2, value=date(int(y), int(m), int(d)))
        ws.cell(row=i, column=2).number_format = 'yyyy-mm-dd'
        ws.cell(row=i, column=3, value=completed)

    # D column (Status) is intentionally left EMPTY -- agent must add formulas
    # E column (Days Remaining) is intentionally left EMPTY -- agent must add formulas
    # G2 (Completion %) is intentionally left EMPTY -- agent must add formula

    # --- Column widths for readability ---
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['G'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
