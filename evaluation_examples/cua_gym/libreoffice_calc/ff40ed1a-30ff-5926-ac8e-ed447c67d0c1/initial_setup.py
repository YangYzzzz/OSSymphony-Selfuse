"""
Initial Setup: NLPSOLVE (Solver for Nonlinear Programming) extension NOT installed
Task ID: osworld_multi_apps_ext_install_009
Domain: libreoffice_calc (multi-app: LibreOffice Calc + Chrome extension install)

Prepares initial state:
- LibreOffice Calc is open with a sample spreadsheet
- The NLPSolver extension is NOT installed
- Chrome is available for the agent to download the extension
"""

import os
import io
import zipfile
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_ext_install_009'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

# The NLPSolver extension identifier in LibreOffice
EXTENSION_ID = 'com.sun.star.comp.solver.NLPSolver'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env['DISPLAY'] = ':0'
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def kill_libreoffice():
    """Kill any running LibreOffice processes."""
    subprocess.run(['pkill', '-f', 'soffice'], capture_output=True)
    time.sleep(2)
    print('LibreOffice processes killed')


def ensure_extension_not_installed():
    """Ensure the NLPSolver extension is NOT installed."""
    result = subprocess.run(
        ['unopkg', 'list'],
        capture_output=True, text=True
    )
    if EXTENSION_ID in result.stdout:
        print(f'Found existing extension {EXTENSION_ID}, removing...')
        remove_result = subprocess.run(
            ['unopkg', 'remove', EXTENSION_ID],
            capture_output=True, text=True
        )
        if remove_result.returncode == 0:
            print('Extension removed successfully')
        else:
            print(f'Warning: could not remove extension: {remove_result.stderr}')
    else:
        print(f'Extension {EXTENSION_ID} not installed (as expected)')


def create_initial_spreadsheet():
    """Create a sample spreadsheet with optimization data for the task context."""
    wb = openpyxl.Workbook()

    # --- Sheet 1: Production Optimization ---
    ws1 = wb.active
    ws1.title = 'Production'

    # Header row with styling
    headers = ['Product', 'Units', 'Profit/Unit', 'Material Cost', 'Labor Hours', 'Total Profit']
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Product data (realistic manufacturing scenario)
    products = [
        ['Widget A',    120, 45.50,  18.20, 2.5,  '=B2*C2'],
        ['Widget B',    85,  62.00,  24.80, 3.2,  '=B3*C3'],
        ['Gadget X',    200, 28.75,  11.40, 1.8,  '=B4*C4'],
        ['Gadget Y',    60,  95.25,  38.60, 4.1,  '=B5*C5'],
        ['Component Z', 340, 15.90,   6.30, 0.9,  '=B6*C6'],
        ['Assembly P',  45,  128.00, 52.40, 5.5,  '=B7*C7'],
        ['Module Q',    90,  74.50,  30.10, 3.8,  '=B8*C8'],
        ['Unit R',      175, 38.25,  15.50, 2.1,  '=B9*C9'],
    ]
    for r, row_data in enumerate(products, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Summary row
    ws1.cell(row=11, column=1, value='TOTAL')
    ws1.cell(row=11, column=1).font = Font(bold=True)
    ws1.cell(row=11, column=6, value='=SUM(F2:F9)')
    ws1.cell(row=11, column=6).font = Font(bold=True)

    # Set column widths
    ws1.column_dimensions['A'].width = 18
    ws1.column_dimensions['B'].width = 10
    ws1.column_dimensions['C'].width = 14
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 14
    ws1.column_dimensions['F'].width = 14

    # Freeze the header row
    ws1.freeze_panes = 'A2'

    # --- Sheet 2: Constraints ---
    ws2 = wb.create_sheet('Constraints')
    ws2.cell(row=1, column=1, value='Constraint').font = Font(bold=True)
    ws2.cell(row=1, column=2, value='Limit').font = Font(bold=True)
    ws2.cell(row=1, column=3, value='Current').font = Font(bold=True)

    constraints = [
        ['Max Material Budget ($)',  25000, '=Production!D2*Production!B2+Production!D3*Production!B3'],
        ['Max Labor Hours/Day',      480,   '=Production!E2*Production!B2+Production!E3*Production!B3'],
        ['Min Widget A Production',  100,   '=Production!B2'],
        ['Max Total Units',          1500,  '=SUM(Production!B2:B9)'],
    ]
    for r, row_data in enumerate(constraints, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 16

    wb.save(OUTPUT)
    print(f'Initial spreadsheet created: {OUTPUT}')


def setup_initial():
    """Set up initial state: LibreOffice Calc open, NLPSolver NOT installed."""
    print('Setting up initial state...')

    # Kill LibreOffice first
    kill_libreoffice()

    # Ensure NLPSolver extension is NOT installed
    ensure_extension_not_installed()

    # Create the initial spreadsheet
    create_initial_spreadsheet()

    print('Initial state ready:')
    print(f'  - Extension {EXTENSION_ID} is NOT installed')
    print(f'  - Spreadsheet available at: {OUTPUT}')

    # Launch LibreOffice Calc with the spreadsheet
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=3.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


setup_initial()
