"""
Initial Setup: Hide rows with blank or N/A Department in employee spreadsheet
Task ID: osworld_calc_hide_rows_na_003
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_hide_rows_na_003'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Employee Data ---
    ws = wb.active
    ws.title = 'Employee Data'

    # Headers
    headers = ['Employee ID', 'Name', 'Department', 'Job Title', 'Salary']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Employee data rows
    # Rows with blank ('') or 'N/A' in Department column are the ones to be hidden
    # Row indices (1-based) with blank/N/A: rows 4, 7, 10, 13, 16
    data = [
        # (Employee ID, Name, Department, Job Title, Salary)
        ('EMP001', 'Sarah Chen',       'Engineering',  'Senior Software Engineer',  95000),
        ('EMP002', 'Marcus Johnson',   'Marketing',    'Marketing Manager',         78000),
        ('EMP003', 'Priya Nair',       'Finance',      'Financial Analyst',          67000),
        ('EMP004', 'Derek Walsh',      'N/A',          'TBD',                        55000),  # row 5 - target
        ('EMP005', 'Linda Okafor',     'HR',           'HR Business Partner',        72000),
        ('EMP006', 'James Thornton',   'Engineering',  'DevOps Engineer',            88000),
        ('EMP007', 'Anika Sharma',     '',             'Contractor',                 48000),  # row 8 - target (blank)
        ('EMP008', 'Robert Castillo',  'Operations',   'Operations Manager',         81000),
        ('EMP009', 'Mei-Lin Huang',    'Finance',      'Senior Accountant',          74000),
        ('EMP010', 'Tyler Brooks',     'N/A',          'Pending Assignment',          52000),  # row 11 - target
        ('EMP011', 'Fatima Al-Hassan', 'Marketing',    'Content Strategist',         65000),
        ('EMP012', 'Nathan Pierce',    'Engineering',  'Backend Developer',          91000),
        ('EMP013', 'Isabelle Martin',  '',             'Intern',                     35000),  # row 14 - target (blank)
        ('EMP014', 'Kwame Asante',     'Operations',   'Supply Chain Analyst',       70000),
        ('EMP015', 'Clara Novak',      'HR',           'Talent Acquisition Lead',    69000),
        ('EMP016', 'Raj Patel',        'N/A',          'Awaiting Onboarding',        58000),  # row 17 - target
        ('EMP017', 'Stephanie Wolfe',  'Finance',      'Controller',                 87000),
        ('EMP018', 'Carlos Mendez',    'Engineering',  'Data Engineer',              84000),
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 12

    # Row height for header
    ws.row_dimensions[1].height = 20

    # NOTE: All rows are VISIBLE in the initial state - the task is to hide the N/A/blank rows
    # Do NOT hide any rows here

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
