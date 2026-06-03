"""
Reward Script: Hide all rows where Department is blank or N/A
Task ID: osworld_calc_hide_rows_na_003
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5 pts): All rows with Department='N/A' are hidden (rows 5, 11, 17)
  Component 2 (0.3 pts): All rows with blank/None Department are hidden (rows 8, 14)
  Component 3 (0.2 pts): Rows with valid departments are NOT hidden — only scored when
                          at least one N/A or blank row has been confirmed hidden (anchors
                          this check to task-introduced changes)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_hide_rows_na_003'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Task: Hide all rows where Department (column C) is blank or 'N/A'.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get the active sheet (Employee Data)
    try:
        ws = wb.active
        print(f"INFO: Active sheet is '{ws.title}' with {ws.max_row} rows, {ws.max_column} columns")
    except Exception as e:
        print(f"CRITICAL: Cannot access worksheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: verify the expected structure (header row, department column at col 3)
    try:
        header_dept = ws.cell(row=1, column=3).value
        if str(header_dept).strip() != 'Department':
            print(f"PRECONDITION FAIL: Column 3 header is '{header_dept}', expected 'Department'")
            print("REWARD: 0.0")
            return 0.0
        print(f"INFO: Precondition check passed — column 3 header is '{header_dept}'")
    except Exception as e:
        print(f"CRITICAL: Cannot verify header structure: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Scan all data rows and categorize them
    # Rows with Department = 'N/A' (exact string) => must be hidden
    # Rows with Department = None/blank               => must be hidden
    # Rows with valid departments                     => must remain visible
    na_rows = []          # rows with 'N/A'
    blank_rows = []       # rows with None/blank
    valid_rows = []       # rows with real department names

    try:
        for row_num in range(2, ws.max_row + 1):
            dept_val = ws.cell(row=row_num, column=3).value
            if dept_val is None or str(dept_val).strip() == '':
                blank_rows.append(row_num)
            elif str(dept_val).strip() == 'N/A':
                na_rows.append(row_num)
            else:
                valid_rows.append(row_num)

        print(f"INFO: N/A department rows: {na_rows}")
        print(f"INFO: Blank/None department rows: {blank_rows}")
        print(f"INFO: Valid department rows: {valid_rows}")
    except Exception as e:
        print(f"CRITICAL: Error scanning rows: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: All N/A rows are hidden (0.5 points)
    # These rows must have been explicitly hidden by the task action.
    # Expected: rows 5, 11, 17 (Department = 'N/A')
    try:
        if len(na_rows) == 0:
            print("WARN: No N/A rows found in the file — cannot verify Component 1")
        else:
            all_na_hidden = all(ws.row_dimensions[r].hidden for r in na_rows)
            hidden_na = [r for r in na_rows if ws.row_dimensions[r].hidden]
            not_hidden_na = [r for r in na_rows if not ws.row_dimensions[r].hidden]

            if all_na_hidden:
                print(f"PASS: Component 1 — All N/A rows are hidden: {hidden_na} (0.5 pts)")
                total_score += 0.5
            else:
                print(f"FAIL: Component 1 — Some N/A rows are NOT hidden. "
                      f"Hidden: {hidden_na}, Not hidden: {not_hidden_na}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All blank/None Department rows are hidden (0.3 points)
    # These rows must also be hidden since blank = no department assigned.
    # Expected: rows 8, 14 (Department = None/blank)
    try:
        if len(blank_rows) == 0:
            print("WARN: No blank department rows found in the file — cannot verify Component 2")
        else:
            all_blank_hidden = all(ws.row_dimensions[r].hidden for r in blank_rows)
            hidden_blank = [r for r in blank_rows if ws.row_dimensions[r].hidden]
            not_hidden_blank = [r for r in blank_rows if not ws.row_dimensions[r].hidden]

            if all_blank_hidden:
                print(f"PASS: Component 2 — All blank/None department rows are hidden: "
                      f"{hidden_blank} (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — Some blank/None rows are NOT hidden. "
                      f"Hidden: {hidden_blank}, Not hidden: {not_hidden_blank}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Rows with valid departments are NOT hidden (0.2 points)
    # This check is ONLY scored when at least one N/A or blank row is confirmed hidden,
    # anchoring it to task-introduced changes (not a precondition on the initial state).
    try:
        any_target_hidden = any(ws.row_dimensions[r].hidden for r in na_rows + blank_rows)
        if not any_target_hidden:
            print("SKIP: Component 3 — No target rows were hidden; skipping collateral-damage check "
                  "(this check is only scored when the task has been partially performed)")
        elif len(valid_rows) == 0:
            print("WARN: No valid department rows found — cannot verify Component 3")
        else:
            all_valid_visible = all(not ws.row_dimensions[r].hidden for r in valid_rows)
            hidden_valid = [r for r in valid_rows if ws.row_dimensions[r].hidden]
            visible_valid = [r for r in valid_rows if not ws.row_dimensions[r].hidden]

            if all_valid_visible:
                print(f"PASS: Component 3 — All valid-department rows remain visible: "
                      f"{len(visible_valid)} rows visible, none incorrectly hidden (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 3 — Some valid-department rows were incorrectly hidden: "
                      f"{hidden_valid}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point: test against canonical artifact path on the VM
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
