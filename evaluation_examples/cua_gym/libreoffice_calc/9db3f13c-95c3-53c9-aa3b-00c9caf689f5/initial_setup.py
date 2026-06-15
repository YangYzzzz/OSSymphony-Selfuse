"""
Initial Setup: Trial balance spreadsheet with raw data, no formatting.
Task ID: calc_gpm_009
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_009'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'TrialBal'

    # Headers (plain, no formatting)
    ws['A1'] = 'Account'
    ws['B1'] = 'Debit'
    ws['C1'] = 'Credit'

    # Data rows
    data = [
        ('Cash', 54000, None),
        ('Accounts Receivable', 32000, None),
        ('Supplies', 4500, None),
        ('Equipment', 85000, None),
        ('Accounts Payable', None, 21500),
        ('Notes Payable', None, 50000),
        ('Common Stock', None, 75000),
        ('Retained Earnings', None, 18000),
        ('Service Revenue', None, 42000),
        ('Salary Expense', 22000, None),
        ('Rent Expense', 9000, None),
    ]

    for r, (account, debit, credit) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=account)
        if debit is not None:
            ws.cell(row=r, column=2, value=debit)
        if credit is not None:
            ws.cell(row=r, column=3, value=credit)

    # Row 13 is empty (separator)
    # Totals row (plain values, no formulas, no formatting)
    ws['A14'] = 'Totals'
    ws['B14'] = 206500
    ws['C14'] = 206500

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
