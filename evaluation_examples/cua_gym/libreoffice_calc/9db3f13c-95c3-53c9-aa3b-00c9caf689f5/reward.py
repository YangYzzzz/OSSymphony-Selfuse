"""
Reward Script: Apply currency formatting and accounting number format to a trial balance spreadsheet.
Task ID: calc_gpm_009
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Header row formatting (bold, centered, gray fill, double bottom border)
  Component 2 (0.30): Accounting number format on B2:C12
  Component 3 (0.20): Totals row formatting (B14:C14 bold, SUM formulas, accounting format)
  Component 4 (0.10): Totals row borders (B14:C14 double top and double bottom border)
  Component 5 (0.15): Column widths (A=25, B=15, C=15)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_009'
ACCOUNTING_FMT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'TrialBal' not in wb.sheetnames:
        print("CRITICAL: Sheet 'TrialBal' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['TrialBal']

    # Component 1: Header row formatting — bold, centered, gray fill, double bottom border (0.25 points)
    # In initial_env: row 1 has no bold, no center, no fill, no border → should FAIL
    try:
        header_checks_passed = 0
        total_header_checks = 4  # bold, centered, fill, border

        # 1a: All three header cells bold
        all_bold = all(ws.cell(row=1, column=c).font.bold for c in range(1, 4))
        if all_bold:
            header_checks_passed += 1
            print("PASS: Component 1a — Header cells A1:C1 are bold")
        else:
            print("FAIL: Component 1a — Not all header cells are bold")

        # 1b: All three header cells centered
        all_centered = all(
            ws.cell(row=1, column=c).alignment.horizontal == 'center'
            for c in range(1, 4)
        )
        if all_centered:
            header_checks_passed += 1
            print("PASS: Component 1b — Header cells A1:C1 are centered")
        else:
            print("FAIL: Component 1b — Not all header cells are centered")

        # 1c: Gray fill (RGB ~200,200,200 → ARGB FFC8C8C8)
        fill_ok = True
        for c in range(1, 4):
            cell = ws.cell(row=1, column=c)
            try:
                fg = cell.fill.fgColor.rgb
                if fg is None or cell.fill.fill_type != 'solid':
                    fill_ok = False
                    break
                # Accept a range around C8C8C8 — allow some tolerance
                r_val = int(fg[2:4], 16)
                g_val = int(fg[4:6], 16)
                b_val = int(fg[6:8], 16)
                if not (180 <= r_val <= 220 and 180 <= g_val <= 220 and 180 <= b_val <= 220):
                    fill_ok = False
                    break
            except Exception:
                fill_ok = False
                break
        if fill_ok:
            header_checks_passed += 1
            print("PASS: Component 1c — Header cells have gray fill")
        else:
            print("FAIL: Component 1c — Header cells missing gray fill")

        # 1d: Double bottom border on row 1
        border_ok = True
        for c in range(1, 4):
            cell = ws.cell(row=1, column=c)
            bot = cell.border.bottom
            if bot is None or bot.style != 'double':
                border_ok = False
                break
        if border_ok:
            header_checks_passed += 1
            print("PASS: Component 1d — Header cells have double bottom border")
        else:
            print("FAIL: Component 1d — Header cells missing double bottom border")

        # Award proportional credit for header formatting
        if header_checks_passed > 0:
            comp1_score = 0.25 * (header_checks_passed / total_header_checks)
            print(f"  Component 1 subtotal: {header_checks_passed}/{total_header_checks} checks passed ({comp1_score:.3f} pts)")
            total_score += comp1_score
        else:
            print("  Component 1 subtotal: 0/4 checks passed (0.0 pts)")

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Accounting number format on B2:C12 (0.30 points)
    # In initial_env: all cells have 'General' format → should FAIL
    try:
        acct_fmt_count = 0
        total_cells_to_check = 22  # B2:C12 = 11 rows * 2 cols

        for r in range(2, 13):
            for c in [2, 3]:
                cell = ws.cell(row=r, column=c)
                nf = cell.number_format
                # Check for accounting format — match the key pattern
                if nf and '_($*' in nf and '#,##0.00' in nf:
                    acct_fmt_count += 1

        if acct_fmt_count == total_cells_to_check:
            print(f"PASS: Component 2 — All {total_cells_to_check} cells in B2:C12 have accounting format (0.30 pts)")
            total_score += 0.30
        elif acct_fmt_count > 0:
            # Partial credit based on proportion
            partial = 0.30 * (acct_fmt_count / total_cells_to_check)
            print(f"PARTIAL: Component 2 — {acct_fmt_count}/{total_cells_to_check} cells have accounting format ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No cells in B2:C12 have accounting format (0/{total_cells_to_check})")

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Totals row formatting — A14 bold, B14:C14 bold + SUM formulas + accounting format (0.20 points)
    # In initial_env: A14 not bold, B14/C14 not bold, values are hardcoded 206500, format is General → should FAIL
    try:
        totals_checks_passed = 0
        total_totals_checks = 4

        # 3a: A14 is bold
        if ws['A14'].font.bold:
            totals_checks_passed += 1
            print("PASS: Component 3a — A14 is bold")
        else:
            print("FAIL: Component 3a — A14 is not bold")

        # 3b: B14 and C14 are bold
        if ws['B14'].font.bold and ws['C14'].font.bold:
            totals_checks_passed += 1
            print("PASS: Component 3b — B14:C14 are bold")
        else:
            print(f"FAIL: Component 3b — B14 bold={ws['B14'].font.bold}, C14 bold={ws['C14'].font.bold}")

        # 3c: B14 and C14 contain SUM formulas
        b14_val = ws['B14'].value
        c14_val = ws['C14'].value
        b14_is_sum = isinstance(b14_val, str) and '=SUM' in b14_val.upper()
        c14_is_sum = isinstance(c14_val, str) and '=SUM' in c14_val.upper()
        if b14_is_sum and c14_is_sum:
            totals_checks_passed += 1
            print(f"PASS: Component 3c — B14={b14_val}, C14={c14_val} (SUM formulas)")
        else:
            print(f"FAIL: Component 3c — B14={b14_val!r}, C14={c14_val!r} (expected SUM formulas)")

        # 3d: B14 and C14 have accounting format
        b14_fmt = ws['B14'].number_format
        c14_fmt = ws['C14'].number_format
        b14_acct = b14_fmt and '_($*' in b14_fmt and '#,##0.00' in b14_fmt
        c14_acct = c14_fmt and '_($*' in c14_fmt and '#,##0.00' in c14_fmt
        if b14_acct and c14_acct:
            totals_checks_passed += 1
            print("PASS: Component 3d — B14:C14 have accounting format")
        else:
            print(f"FAIL: Component 3d — B14 fmt={b14_fmt!r}, C14 fmt={c14_fmt!r}")

        if totals_checks_passed > 0:
            comp3_score = 0.20 * (totals_checks_passed / total_totals_checks)
            print(f"  Component 3 subtotal: {totals_checks_passed}/{total_totals_checks} checks passed ({comp3_score:.3f} pts)")
            total_score += comp3_score
        else:
            print("  Component 3 subtotal: 0/4 checks passed (0.0 pts)")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Totals row borders — B14:C14 double top and double bottom border (0.10 points)
    # In initial_env: no borders → should FAIL
    try:
        borders_ok = True
        for c in [2, 3]:
            cell = ws.cell(row=14, column=c)
            top = cell.border.top
            bot = cell.border.bottom
            if top is None or top.style != 'double':
                borders_ok = False
                print(f"FAIL: Component 4 — {cell.coordinate} top border style={top.style if top else None}, expected 'double'")
                break
            if bot is None or bot.style != 'double':
                borders_ok = False
                print(f"FAIL: Component 4 — {cell.coordinate} bottom border style={bot.style if bot else None}, expected 'double'")
                break
        if borders_ok:
            print("PASS: Component 4 — B14:C14 have double top and double bottom borders (0.10 pts)")
            total_score += 0.10
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Column widths — A=25, B=15, C=15 (0.15 points)
    # In initial_env: all columns are 13.0 → should FAIL
    try:
        width_checks = 0
        total_width_checks = 3

        col_a_w = ws.column_dimensions['A'].width
        col_b_w = ws.column_dimensions['B'].width
        col_c_w = ws.column_dimensions['C'].width

        # Allow tolerance of 1 unit to avoid scoring pre-existing widths
        if col_a_w is not None and abs(col_a_w - 25) <= 1:
            width_checks += 1
            print(f"PASS: Component 5a — Column A width={col_a_w} (expected ~25)")
        else:
            print(f"FAIL: Component 5a — Column A width={col_a_w} (expected ~25)")

        if col_b_w is not None and abs(col_b_w - 15) <= 1:
            width_checks += 1
            print(f"PASS: Component 5b — Column B width={col_b_w} (expected ~15)")
        else:
            print(f"FAIL: Component 5b — Column B width={col_b_w} (expected ~15)")

        if col_c_w is not None and abs(col_c_w - 15) <= 1:
            width_checks += 1
            print(f"PASS: Component 5c — Column C width={col_c_w} (expected ~15)")
        else:
            print(f"FAIL: Component 5c — Column C width={col_c_w} (expected ~15)")

        if width_checks > 0:
            comp5_score = 0.15 * (width_checks / total_width_checks)
            print(f"  Component 5 subtotal: {width_checks}/{total_width_checks} checks passed ({comp5_score:.3f} pts)")
            total_score += comp5_score
        else:
            print("  Component 5 subtotal: 0/3 checks passed (0.0 pts)")

    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
