"""
Initial Setup: ACL Conference Locations spreadsheet (Host City column empty)
Task ID: osworld_multi_apps_conference_city_008
Domain: libreoffice_calc

Creates ACL_Conferences.xlsx with Year (2015-2022) and empty Host City column.
The agent must search the web to fill in the Host City for each year.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_conference_city_008'
OUTPUT = f'{WORKDIR}/ACL_Conferences.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ACL Conferences"

    # --- Header row ---
    header_font = Font(name="Calibri", size=12, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Year", "Host City"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # --- Data rows: Years 2015-2022, Host City BLANK ---
    years = [2015, 2016, 2017, 2018, 2019, 2020, 2021, 2022]
    data_align = Alignment(horizontal="center", vertical="center")
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, year in enumerate(years, 2):
        # Year column
        year_cell = ws.cell(row=row_idx, column=1, value=year)
        year_cell.font = Font(name="Calibri", size=11)
        year_cell.alignment = data_align
        year_cell.border = data_border
        year_cell.number_format = '0'

        # Host City column — intentionally LEFT BLANK (agent must fill this)
        city_cell = ws.cell(row=row_idx, column=2, value=None)
        city_cell.font = Font(name="Calibri", size=11)
        city_cell.alignment = data_align
        city_cell.border = data_border

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30

    # Row heights
    ws.row_dimensions[1].height = 22
    for r in range(2, 10):
        ws.row_dimensions[r].height = 18

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
