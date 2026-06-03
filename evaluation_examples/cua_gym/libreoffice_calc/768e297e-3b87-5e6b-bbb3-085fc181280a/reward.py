"""
Reward Script: Fill in ACL Conference host cities for years 2015-2022
Task ID: osworld_multi_apps_conference_city_008
Domain: libreoffice_calc
Scoring:
  - Component 1: All 8 Host City cells are filled (non-empty) — 0.4 points (0.05 per city)
  - Component 2: City values match known correct ACL conference locations — 0.6 points (0.075 per city)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_conference_city_008'

# Ground truth ACL conference host cities (keyword-based, case-insensitive substring match)
EXPECTED_CITIES = {
    2015: ['beijing'],
    2016: ['berlin'],
    2017: ['vancouver'],
    2018: ['melbourne'],
    2019: ['florence'],
    2020: ['online'],
    2021: ['bangkok', 'online'],  # 2021 was hybrid (Bangkok/Online)
    2022: ['dublin'],
}

# Years expected in the spreadsheet
YEARS_ORDER = [2015, 2016, 2017, 2018, 2019, 2020, 2021, 2022]


def normalize(val):
    """Normalize a city value for comparison."""
    if val is None:
        return ''
    return str(val).strip().lower()


def city_matches(year, val):
    """Check if a city value matches any expected keyword for the year."""
    normalized = normalize(val)
    if not normalized:
        return False
    for keyword in EXPECTED_CITIES.get(year, []):
        if keyword in normalized:
            return True
    return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Locate the active/only sheet
    try:
        ws = wb.active
        print(f"Sheet: {ws.title}, Rows: {ws.max_row}, Cols: {ws.max_column}")
    except Exception as e:
        print(f"CRITICAL: Cannot access sheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Parse the year -> host city mapping from the spreadsheet
    # Expected layout: Row 1 = headers (Year, Host City), rows 2-9 = data
    year_to_city = {}
    try:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
            year_cell = row[0]
            city_cell = row[1]
            if year_cell.value is not None:
                try:
                    year = int(year_cell.value)
                    year_to_city[year] = city_cell.value
                except (ValueError, TypeError):
                    pass
    except Exception as e:
        print(f"CRITICAL: Cannot read data rows: {e}")
        print("REWARD: 0.0")
        return 0.0

    print(f"Parsed year->city mapping: {year_to_city}")

    # Component 1: All 8 Host City cells are non-empty (0.4 points, 0.05 per city)
    # This FAILS on initial (all None) and PASSES on golden (all filled)
    for year in YEARS_ORDER:
        try:
            city_val = year_to_city.get(year)
            if city_val is not None and str(city_val).strip() != '':
                print(f"PASS: Component 1 — Year {year} has city value: '{city_val}' (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 1 — Year {year} has no city value (found: {city_val!r})")
        except Exception as e:
            print(f"ERROR: Component 1 year {year} — {e}")

    print(f"Component 1 subtotal after 8 years: {total_score:.3f} pts")

    # Component 2: City values match known correct ACL locations (0.6 points, 0.075 per city)
    # This FAILS on initial (all None) and PASSES on golden (correct cities)
    comp2_start = total_score
    for year in YEARS_ORDER:
        try:
            city_val = year_to_city.get(year)
            if city_matches(year, city_val):
                print(f"PASS: Component 2 — Year {year} city '{city_val}' matches expected {EXPECTED_CITIES[year]} (0.075 pts)")
                total_score += 0.075
            else:
                print(f"FAIL: Component 2 — Year {year} city '{city_val}' does not match expected {EXPECTED_CITIES[year]}")
        except Exception as e:
            print(f"ERROR: Component 2 year {year} — {e}")

    print(f"Component 2 subtotal: {total_score - comp2_start:.3f} pts")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score:.4f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the VM
file_path = f'{WORKDIR}/ACL_Conferences.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
