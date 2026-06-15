"""
Initial Setup: Configure data validation error alert on cell C2
Task ID: calc_nrv_053
Domain: libreoffice_calc

Creates a spreadsheet with student scores. Cell C2 has whole number validation
(0-100) but uses the default error alert (no custom title or message).
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_053'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Scores ---
    ws = wb.active
    ws.title = 'Scores'

    # Headers
    headers = ['Student Name', 'Subject', 'Score']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Student data - realistic content
    data = [
        ['Sarah Chen', 'Mathematics', 92],
        ['Marcus Johnson', 'Mathematics', 78],
        ['Priya Patel', 'Mathematics', 85],
        ['James O\'Brien', 'Mathematics', 63],
        ['Aisha Rahman', 'Mathematics', 91],
        ['Carlos Mendez', 'Mathematics', 74],
        ['Emily Watson', 'Mathematics', 88],
        ['Liam Nakamura', 'Mathematics', 56],
        ['Sophie Laurent', 'Mathematics', 95],
        ['David Kim', 'Mathematics', 81],
        ['Olivia Torres', 'Mathematics', 72],
        ['Ryan Mitchell', 'Mathematics', 67],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 12

    # Add whole number validation (0-100) on C2:C13 with DEFAULT error alert
    # No custom error title or message - just the basic validation
    dv = DataValidation(
        type='whole',
        operator='between',
        formula1='0',
        formula2='100',
        allow_blank=True,
    )
    # Explicitly leave error alert as default (no custom title/message)
    dv.showErrorMessage = True
    dv.errorStyle = None  # default
    dv.error = None
    dv.errorTitle = None
    dv.add('C2:C13')
    ws.add_data_validation(dv)

    # --- Sheet 2: Summary ---
    ws2 = wb.create_sheet('Summary')
    ws2['A1'] = 'Class Statistics'
    ws2['A1'].font = Font(bold=True, size=13)
    ws2['A3'] = 'Total Students'
    ws2['B3'] = 12
    ws2['A4'] = 'Subject'
    ws2['B4'] = 'Mathematics'
    ws2['A5'] = 'Semester'
    ws2['B5'] = 'Spring 2025'
    ws2['A6'] = 'Instructor'
    ws2['B6'] = 'Dr. Rebecca Foster'
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 20

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
