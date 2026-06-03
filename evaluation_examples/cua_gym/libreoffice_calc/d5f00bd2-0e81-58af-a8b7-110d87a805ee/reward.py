"""
Reward Script: Fill growth rate formula in column E and create formatted report strings in column F
Task ID: osworld_calc_formula_pattern_concat_009
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): Column E has growth rate formulas =(Cn-Dn)/Dn*100 for rows 3-11
                     (E2 already existed in initial; rows 3-11 are new)
  Component 2 (0.5): Column F has formatted report string formulas for all rows 2-11
                     using TEXT() and & concatenation
Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_formula_pattern_concat_009'


def normalize_formula(f):
    """Normalize formula for comparison: uppercase, remove spaces."""
    if f is None:
        return ''
    return str(f).upper().replace(' ', '')


def is_growth_rate_formula(formula, row):
    """
    Check if the formula matches the growth rate pattern:
    =(Cn-Dn)/Dn*100  for row n
    Accepts both = and case-insensitive variants.
    """
    if not formula:
        return False
    # Normalize
    norm = normalize_formula(formula)
    # Expected pattern: =(C{row}-D{row})/D{row}*100
    expected = f'=(C{row}-D{row})/D{row}*100'
    return norm == normalize_formula(expected)


def has_f_formula_structure(formula, row):
    """
    Check if column F formula for the given row has the required structure:
    - References A{row}, B{row}, C{row}, E{row}
    - Uses TEXT() function for C and E values
    - Contains the pattern: year, metric name, $value, (Growth: rate%)
    Accepts any valid formula that matches the required pattern.
    """
    if not formula:
        return False
    if not isinstance(formula, str) or not formula.startswith('='):
        return False

    # Normalize for keyword checks (case-insensitive, space-free)
    norm = normalize_formula(formula)
    # Original formula for dollar-sign check (spaces matter in string literals)
    formula_upper = formula.upper()

    # Must reference year column (A), metric name (B), value (C), growth rate (E)
    has_year = f'A{row}' in norm
    has_metric = f'B{row}' in norm
    has_value = f'C{row}' in norm
    has_growth = f'E{row}' in norm

    # Must use TEXT() function for formatting
    has_text = 'TEXT(' in norm

    # Must include dollar sign (either as ": $" or ":$" after normalization)
    has_dollar = '$' in formula_upper

    # Must include "growth" label
    has_growth_label = 'GROWTH' in norm

    return has_year and has_metric and has_value and has_growth and has_text and has_dollar and has_growth_label


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task:
    1. Fill column E with growth rate formula =(Cn-Dn)/Dn*100 for rows 3-11
       (E2 already had the formula in the initial state; rows 3-11 are new)
    2. Fill column F with formatted report string formulas for all rows 2-11
       using year, metric name, TEXT(value, "0.00"), and TEXT(growth, "0.00")
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet exists
    if 'Financial Report' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Financial Report' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Financial Report']

    # Precondition gate: check that data rows exist (rows 2-11)
    if ws.max_row < 11:
        print(f"CRITICAL: Expected at least 11 rows, found {ws.max_row}")
        print("REWARD: 0.0")
        return 0.0

    # -----------------------------------------------------------------------
    # Component 1: Column E is filled with growth rate formulas for rows 3-11
    # (0.5 points)
    # In the initial state, only E2 has the formula. Rows 3-11 are empty.
    # The task requires filling the formula down for all remaining rows.
    # -----------------------------------------------------------------------
    try:
        e_filled_count = 0
        e_total_expected = 9  # rows 3 through 11
        e_failures = []

        for row in range(3, 12):  # rows 3 to 11 (new formulas, not present in initial)
            cell_val = ws.cell(row=row, column=5).value  # column E
            if cell_val is not None and isinstance(cell_val, str) and cell_val.startswith('='):
                # Verify it's a growth rate formula
                if is_growth_rate_formula(cell_val, row):
                    e_filled_count += 1
                else:
                    e_failures.append(f"E{row}: formula does not match growth rate pattern: {cell_val}")
            else:
                e_failures.append(f"E{row}: no formula found (value: {repr(cell_val)})")

        if e_filled_count == e_total_expected:
            print(f"PASS: Component 1 — Column E growth rate formulas filled for rows 3-11 ({e_filled_count}/{e_total_expected}) (0.5 pts)")
            total_score += 0.5
        elif e_filled_count > 0:
            partial = round(0.5 * (e_filled_count / e_total_expected), 4)
            total_score += partial
            print(f"PARTIAL: Component 1 — Column E: {e_filled_count}/{e_total_expected} growth rate formulas filled ({partial:.4f} pts)")
            for f in e_failures[:3]:
                print(f"  - {f}")
        else:
            print(f"FAIL: Component 1 — Column E has no new growth rate formulas in rows 3-11 (0.0 pts)")
            for f in e_failures[:3]:
                print(f"  - {f}")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not check column E: {e}")

    # -----------------------------------------------------------------------
    # Component 2: Column F has formatted report string formulas for rows 2-11
    # (0.5 points)
    # In the initial state, column F is entirely empty.
    # The task requires creating formatted strings in ALL rows 2-11.
    # -----------------------------------------------------------------------
    try:
        f_filled_count = 0
        f_total_expected = 10  # rows 2 through 11
        f_failures = []

        for row in range(2, 12):  # rows 2 to 11
            cell_val = ws.cell(row=row, column=6).value  # column F
            if cell_val is not None and isinstance(cell_val, str) and cell_val.startswith('='):
                # Verify it has the required structure
                if has_f_formula_structure(cell_val, row):
                    f_filled_count += 1
                else:
                    f_failures.append(f"F{row}: formula present but missing required elements: {cell_val[:80]}")
            else:
                f_failures.append(f"F{row}: no formula found (value: {repr(cell_val)})")

        if f_filled_count == f_total_expected:
            print(f"PASS: Component 2 — Column F report string formulas filled for rows 2-11 ({f_filled_count}/{f_total_expected}) (0.5 pts)")
            total_score += 0.5
        elif f_filled_count > 0:
            partial = round(0.5 * (f_filled_count / f_total_expected), 4)
            total_score += partial
            print(f"PARTIAL: Component 2 — Column F: {f_filled_count}/{f_total_expected} report string formulas filled ({partial:.4f} pts)")
            for f in f_failures[:3]:
                print(f"  - {f}")
        else:
            print(f"FAIL: Component 2 — Column F has no report string formulas (0.0 pts)")
            for f in f_failures[:3]:
                print(f"  - {f}")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not check column F: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.4f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
