"""
Initial Setup: Financial report spreadsheet with growth rate formula in E2 only,
               column F empty, and remaining E column empty (to be filled by agent).
Task ID: osworld_calc_formula_pattern_concat_009
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_formula_pattern_concat_009'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Financial Report ---
    ws = wb.active
    ws.title = 'Financial Report'

    # Headers
    headers = ['Year', 'Metric Name', 'Value', 'Prior Year Value', 'Growth Rate %']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # Column F header (empty for now — agent will fill it)
    ws.cell(row=1, column=6, value='Report String')
    cell_f1 = ws.cell(row=1, column=6)
    cell_f1.font = Font(bold=True, color="FFFFFFFF")
    cell_f1.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    cell_f1.alignment = Alignment(horizontal="center")

    # Realistic annual financial data (10 rows, years 2015-2024)
    # (Year, Metric Name, Value, Prior Year Value)
    data = [
        (2015, 'Total Revenue',       4250000.00,  3980000.00),
        (2016, 'Total Revenue',       4612000.00,  4250000.00),
        (2017, 'Total Revenue',       5034000.00,  4612000.00),
        (2018, 'Total Revenue',       5287000.00,  5034000.00),
        (2019, 'Total Revenue',       5841000.00,  5287000.00),
        (2020, 'Total Revenue',       5312000.00,  5841000.00),
        (2021, 'Total Revenue',       6178000.00,  5312000.00),
        (2022, 'Total Revenue',       7045000.00,  6178000.00),
        (2023, 'Total Revenue',       7892000.00,  7045000.00),
        (2024, 'Total Revenue',       8654000.00,  7892000.00),
    ]

    for r, (year, metric, value, prior) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=year)
        ws.cell(row=r, column=2, value=metric)
        ws.cell(row=r, column=3, value=value)
        ws.cell(row=r, column=4, value=prior)
        # Column E: only E2 has the formula; E3:E11 are left empty
        if r == 2:
            ws.cell(row=r, column=5, value='=(C2-D2)/D2*100')
        # Column F: all empty (agent fills this)

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 55

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
