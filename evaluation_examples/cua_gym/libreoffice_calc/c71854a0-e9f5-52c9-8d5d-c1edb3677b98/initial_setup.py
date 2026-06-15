"""
Initial Setup: Add line sparklines in column G for sales trend
Task ID: calc_chart_sparkline_line_046
Domain: libreoffice_calc

Creates the initial spreadsheet with sales data but NO sparklines.
The Trend column (G) header exists but G2:G6 are empty.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_sparkline_line_046'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: SparkData ---
    ws = wb.active
    ws.title = 'SparkData'

    # Headers in row 1
    headers = ['Product', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Trend']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Sales data rows 2-6 (NO sparklines - G2:G6 remain empty)
    data = [
        ['Product A', 1200, 1350, 1280, 1420, 1580],
        ['Product B', 980,  920,  1050, 1100, 1240],
        ['Product C', 2100, 2250, 2180, 2380, 2520],
        ['Product D', 650,  700,  720,  690,  760],
        ['Product E', 1450, 1380, 1500, 1620, 1740],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 1:
                cell.font = Font(name='Calibri', size=11)
            else:
                cell.alignment = Alignment(horizontal='right')
        # G column (col 7) intentionally left empty - no sparklines yet

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 14

    # Row heights
    ws.row_dimensions[1].height = 18
    for r in range(2, 7):
        ws.row_dimensions[r].height = 22

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheet: SparkData')
    print('Rows: 1 header + 5 data rows (Products A-E)')
    print('Columns: Product, Jan, Feb, Mar, Apr, May, Trend (G empty - no sparklines)')


create_initial()
