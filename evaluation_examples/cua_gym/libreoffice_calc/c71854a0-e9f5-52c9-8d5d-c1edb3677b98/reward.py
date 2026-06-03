"""
Reward Script: Add line sparklines in column G to show sales trend for each product
Task ID: calc_chart_sparkline_line_046
Domain: libreoffice_calc
Scoring:
  - Component 1: Sparkline group exists and is of type 'line' (0.4 pts)
  - Component 2: All 5 sparklines present in G2:G6 with correct data references B:F (0.4 pts)
  - Component 3: Original product data in B2:F6 is intact (0.2 pts)
"""

import os
import zipfile
import re

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_chart_sparkline_line_046'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.

    Task: Add line sparklines in column G (G2:G6) for each product's 5-month trend (B:F).

    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: file must exist and be a valid xlsx
    if not os.path.exists(file_path):
        print(f"CRITICAL: File not found: {file_path}")
        print("REWARD: 0.0")
        return 0.0

    try:
        with zipfile.ZipFile(file_path) as z:
            file_list = z.namelist()
            sheet_xml = z.read('xl/worksheets/sheet1.xml').decode('utf-8')
    except Exception as e:
        print(f"CRITICAL: Cannot open xlsx as zip: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Also load with openpyxl for data verification
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
        ws = wb['SparkData']
    except Exception as e:
        print(f"CRITICAL: Cannot load workbook with openpyxl: {e}")
        print("REWARD: 0.0")
        return 0.0

    # ---------------------------------------------------------------------------
    # Component 1: Sparkline group exists and is of type 'line' (0.4 points)
    # The task requires LINE sparklines. The initial file has NO sparklines.
    # This component fails on initial (no sparkline groups), passes on golden (line sparkline group present).
    # ---------------------------------------------------------------------------
    try:
        # Check for sparklineGroups element in the worksheet XML
        has_sparkline_groups = 'x14:sparklineGroups' in sheet_xml or 'sparklineGroups' in sheet_xml

        if not has_sparkline_groups:
            print("FAIL: Component 1 — No sparkline groups found in the worksheet")
        else:
            # Extract sparklineGroup type attribute
            type_match = re.search(r'<x14:sparklineGroup[^>]*type=["\']([^"\']+)["\']', sheet_xml)
            if not type_match:
                # Try alternative namespace
                type_match = re.search(r'<sparklineGroup[^>]*type=["\']([^"\']+)["\']', sheet_xml)

            if type_match:
                sparkline_type = type_match.group(1)
                if sparkline_type == 'line':
                    print(f"PASS: Component 1 — Line sparkline group found (type='{sparkline_type}') (0.4 pts)")
                    total_score += 0.4
                else:
                    print(f"FAIL: Component 1 — Sparkline group found but type='{sparkline_type}', expected 'line'")
            else:
                # If sparklineGroups is present but type not found, it may default to line
                # Check if it at least has sparklines
                if '<x14:sparklines>' in sheet_xml or '<sparklines>' in sheet_xml:
                    print("PASS: Component 1 — Sparkline group found (type defaults to line) (0.4 pts)")
                    total_score += 0.4
                else:
                    print("FAIL: Component 1 — Sparkline group tag found but no sparklines inside")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ---------------------------------------------------------------------------
    # Component 2: All 5 sparklines in G2:G6 with correct data references B:F (0.4 points)
    # Each sparkline should reference the row's 5-month data: SparkData!B{row}:F{row}
    # Initial file has no sparklines, so this always fails on initial.
    # ---------------------------------------------------------------------------
    try:
        # Extract all sparkline definitions
        # Pattern: <x14:sparkline><xm:f>...</xm:f><xm:sqref>...</xm:sqref></x14:sparkline>
        sparkline_pattern = re.compile(
            r'<x14:sparkline>\s*<xm:f>([^<]+)</xm:f>\s*<xm:sqref>([^<]+)</xm:sqref>\s*</x14:sparkline>'
        )
        sparklines = sparkline_pattern.findall(sheet_xml)

        if not sparklines:
            # Try alternative without namespace
            sparkline_pattern2 = re.compile(
                r'<sparkline>\s*<f>([^<]+)</f>\s*<sqref>([^<]+)</sqref>\s*</sparkline>'
            )
            sparklines = sparkline_pattern2.findall(sheet_xml)

        if len(sparklines) == 0:
            print("FAIL: Component 2 — No individual sparklines found in worksheet XML")
        else:
            # Expected: 5 sparklines, each mapping G{row} -> SparkData!B{row}:F{row} for rows 2-6
            expected = {
                'G2': 'SparkData!B2:F2',
                'G3': 'SparkData!B3:F3',
                'G4': 'SparkData!B4:F4',
                'G5': 'SparkData!B5:F5',
                'G6': 'SparkData!B6:F6',
            }

            found_map = {}
            for data_ref, cell_ref in sparklines:
                found_map[cell_ref.strip()] = data_ref.strip()

            correct_count = 0
            issues = []
            for cell, expected_data in expected.items():
                if cell in found_map:
                    actual_data = found_map[cell]
                    if actual_data == expected_data:
                        correct_count += 1
                    else:
                        # Also accept without sheet prefix if sheet is SparkData
                        # e.g. "B2:F2" instead of "SparkData!B2:F2"
                        row_num = cell[1]  # e.g. '2' from 'G2'
                        alt_ref = f'B{row_num}:F{row_num}'
                        if actual_data == alt_ref:
                            correct_count += 1
                        else:
                            issues.append(f"{cell}: expected data ref '{expected_data}', found '{actual_data}'")
                else:
                    issues.append(f"{cell}: sparkline missing")

            if correct_count == 5:
                print(f"PASS: Component 2 — All 5 sparklines present in G2:G6 with correct data references (0.4 pts)")
                total_score += 0.4
            elif correct_count >= 3:
                # Partial credit: most sparklines are correct
                partial = round(0.4 * correct_count / 5, 2)
                print(f"PARTIAL: Component 2 — {correct_count}/5 sparklines correct ({partial} pts)")
                print(f"  Issues: {issues}")
                if correct_count >= 3:
                    total_score += partial
            else:
                print(f"FAIL: Component 2 — Only {correct_count}/5 sparklines correct")
                print(f"  Issues: {issues}")
                print(f"  Found sparklines: {found_map}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---------------------------------------------------------------------------
    # Component 3: Original product data in B2:F6 is preserved (0.2 points)
    # This is a combined check: sparklines exist AND data is intact.
    # Since the task is to ADD sparklines without modifying data, this ensures the agent didn't corrupt data.
    # This is scored as part of the task completion verification - only meaningful when sparklines are also present.
    # ---------------------------------------------------------------------------
    try:
        expected_data = {
            (2, 2): 1200, (2, 3): 1350, (2, 4): 1280, (2, 5): 1420, (2, 6): 1580,
            (3, 2): 980,  (3, 3): 920,  (3, 4): 1050, (3, 5): 1100, (3, 6): 1240,
            (4, 2): 2100, (4, 3): 2250, (4, 4): 2180, (4, 5): 2380, (4, 6): 2520,
            (5, 2): 650,  (5, 3): 700,  (5, 4): 720,  (5, 5): 690,  (5, 6): 760,
            (6, 2): 1450, (6, 3): 1380, (6, 4): 1500, (6, 5): 1620, (6, 6): 1740,
        }

        wrong_cells = []
        from openpyxl.utils import get_column_letter
        for (row, col), exp_val in expected_data.items():
            actual_val = ws.cell(row=row, column=col).value
            try:
                actual_num = int(actual_val) if actual_val is not None else None
            except (ValueError, TypeError):
                actual_num = None

            if actual_num != exp_val:
                cell_ref = f"{get_column_letter(col)}{row}"
                wrong_cells.append(f"{cell_ref}: expected {exp_val}, got {actual_val}")
        data_ok = (len(wrong_cells) == 0)

        # This component only awards points when sparklines are present (i.e., component 1 passed)
        # AND data is intact - this ensures reward(initial) = 0.0 because initial has no sparklines
        has_sparklines = 'x14:sparklineGroups' in sheet_xml or 'sparklineGroups' in sheet_xml
        if data_ok and has_sparklines:
            print(f"PASS: Component 3 — All product data intact and sparklines present (0.2 pts)")
            total_score += 0.2
        elif data_ok and not has_sparklines:
            print(f"FAIL: Component 3 — Data is intact but no sparklines found (0.0 pts)")
        else:
            print(f"FAIL: Component 3 — Data integrity issues: {wrong_cells[:5]}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
