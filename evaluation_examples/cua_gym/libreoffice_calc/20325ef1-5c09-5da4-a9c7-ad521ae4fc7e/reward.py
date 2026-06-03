"""
Reward Script: 2024 Monthly Calendar in LibreOffice Calc
Task ID: calc_grs_012
Domain: libreoffice_calc
Scoring:
  C1: 12 month sheets with correct names (0.25)
  C2: Header row - month+year, merged A1:G1, bold, large, dark bg, white text (0.20)
  C3: Day headers row 2 Sunday-Saturday, bold (0.10)
  C4: DATE formulas for date positioning (0.15)
  C5: Merged cells for date entries (2-row merges) (0.10)
  C6: Conditional formatting on Sunday/Saturday columns (0.10)
  C7: Navigation hyperlinks between months (0.10)
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_012'

MONTH_NAMES = [
    'January', 'February', 'March', 'April', 'May', 'June',
    'July', 'August', 'September', 'October', 'November', 'December'
]


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    sheet_names = wb.sheetnames

    # Component 1: 12 month sheets with correct names (0.25 points)
    try:
        found_months = [m for m in MONTH_NAMES if m in sheet_names]
        month_count = len(found_months)
        if month_count == 12:
            print(f"PASS: Component 1 — All 12 month sheets found (0.25 pts)")
            total_score += 0.25
        elif month_count >= 6:
            partial = round(0.25 * month_count / 12, 3)
            print(f"PARTIAL: Component 1 — {month_count}/12 month sheets found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {month_count}/12 month sheets found: {found_months}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # If fewer than 2 month sheets, nothing else to check meaningfully
    if len([m for m in MONTH_NAMES if m in sheet_names]) < 2:
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {min(total_score, 1.0)}")
        return min(total_score, 1.0)

    # Component 2: Header row - month+year merged A1:G1, bold, large, dark bg, white text (0.20 pts)
    try:
        header_pass_count = 0
        for month in MONTH_NAMES:
            if month not in sheet_names:
                continue
            ws = wb[month]
            a1 = ws.cell(row=1, column=1)
            val = a1.value

            # Check value contains month name and 2024
            if val and month.lower() in str(val).lower() and '2024' in str(val):
                # Check merged A1:G1
                merged_strs = [str(mr) for mr in ws.merged_cells.ranges]
                is_merged = 'A1:G1' in merged_strs
                # Check bold
                is_bold = a1.font.bold == True
                # Check large font (>= 14)
                is_large = a1.font.size is not None and a1.font.size >= 14
                # Check dark background fill
                has_fill = False
                try:
                    fg = a1.fill.fgColor.rgb
                    if fg and fg != '00000000' and a1.fill.fill_type == 'solid':
                        has_fill = True
                except:
                    pass
                # Check white-ish font color
                has_white_font = False
                try:
                    fc = a1.font.color.rgb
                    if fc:
                        # Accept white or near-white (last 6 chars)
                        rgb_part = fc[-6:]
                        r, g, b = int(rgb_part[0:2], 16), int(rgb_part[2:4], 16), int(rgb_part[4:6], 16)
                        if r > 200 and g > 200 and b > 200:
                            has_white_font = True
                except:
                    pass

                if is_merged and is_bold and is_large and has_fill and has_white_font:
                    header_pass_count += 1
                elif is_merged and is_bold and is_large:
                    header_pass_count += 0.5  # partial: structure ok but styling incomplete

        if header_pass_count >= 11:
            print(f"PASS: Component 2 — {header_pass_count}/12 month headers correct (0.20 pts)")
            total_score += 0.20
        elif header_pass_count >= 6:
            partial = round(0.20 * header_pass_count / 12, 3)
            print(f"PARTIAL: Component 2 — {header_pass_count}/12 month headers correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {header_pass_count}/12 month headers correct")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Day headers row 2 (Sunday-Saturday), bold (0.10 pts)
    try:
        expected_days = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
        day_header_ok_count = 0
        for month in MONTH_NAMES:
            if month not in sheet_names:
                continue
            ws = wb[month]
            row2_vals = []
            all_bold = True
            for c in range(1, 8):
                cell = ws.cell(row=2, column=c)
                row2_vals.append(str(cell.value).strip() if cell.value else '')
                if not cell.font.bold:
                    all_bold = False

            # Check day names match (case insensitive)
            days_match = all(
                row2_vals[i].lower() == expected_days[i].lower()
                for i in range(7)
            )
            if days_match and all_bold:
                day_header_ok_count += 1

        if day_header_ok_count >= 11:
            print(f"PASS: Component 3 — {day_header_ok_count}/12 sheets have correct day headers (0.10 pts)")
            total_score += 0.10
        elif day_header_ok_count >= 6:
            partial = round(0.10 * day_header_ok_count / 12, 3)
            print(f"PARTIAL: Component 3 — {day_header_ok_count}/12 (0{partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {day_header_ok_count}/12 sheets have correct day headers")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: DATE formulas for date positioning (0.15 pts)
    try:
        date_formula_count = 0
        for month in MONTH_NAMES:
            if month not in sheet_names:
                continue
            ws = wb[month]
            has_date_formula = False
            for r in range(3, ws.max_row + 1):
                for c in range(1, 8):
                    val = ws.cell(row=r, column=c).value
                    if isinstance(val, str) and '=DATE(' in val.upper():
                        has_date_formula = True
                        break
                if has_date_formula:
                    break
            if has_date_formula:
                date_formula_count += 1

        if date_formula_count >= 11:
            print(f"PASS: Component 4 — {date_formula_count}/12 sheets use DATE formulas (0.15 pts)")
            total_score += 0.15
        elif date_formula_count >= 6:
            partial = round(0.15 * date_formula_count / 12, 3)
            print(f"PARTIAL: Component 4 — {date_formula_count}/12 sheets use DATE formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {date_formula_count}/12 sheets use DATE formulas")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Merged cells for date entries (0.10 pts)
    # Each date cell should span 2 rows (e.g., B3:B4, C3:C4, etc.)
    try:
        merge_ok_count = 0
        for month in MONTH_NAMES:
            if month not in sheet_names:
                continue
            ws = wb[month]
            merged_strs = [str(mr) for mr in ws.merged_cells.ranges]
            # Count 2-row vertical merges in the date area (rows 3+, cols A-G)
            # Exclude the header merge A1:G1
            date_merges = 0
            for mr_str in merged_strs:
                if mr_str == 'A1:G1':
                    continue
                try:
                    # Parse merge range to check if it's a 2-row vertical merge
                    parts = mr_str.split(':')
                    if len(parts) == 2:
                        from openpyxl.utils.cell import coordinate_from_string
                        col1, row1 = coordinate_from_string(parts[0])
                        col2, row2 = coordinate_from_string(parts[1])
                        if col1 == col2 and row2 - row1 == 1 and row1 >= 3:
                            date_merges += 1
                except:
                    pass
            # A month with ~31 days should have many 2-row merges
            if date_merges >= 15:
                merge_ok_count += 1

        if merge_ok_count >= 11:
            print(f"PASS: Component 5 — {merge_ok_count}/12 sheets have merged date cells (0.10 pts)")
            total_score += 0.10
        elif merge_ok_count >= 6:
            partial = round(0.10 * merge_ok_count / 12, 3)
            print(f"PARTIAL: Component 5 — {merge_ok_count}/12 sheets ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — Only {merge_ok_count}/12 sheets have sufficient merged date cells")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Conditional formatting on Sunday (col A) and Saturday (col G) columns (0.10 pts)
    try:
        cf_ok_count = 0
        for month in MONTH_NAMES:
            if month not in sheet_names:
                continue
            ws = wb[month]
            # Check for conditional formatting rules that cover column A or column G
            has_col_a_cf = False
            has_col_g_cf = False
            for cf in ws.conditional_formatting:
                cf_range = str(cf)
                # Check if the CF range covers column A or G in the date area
                if 'A' in cf_range and any(c.isdigit() for c in cf_range):
                    has_col_a_cf = True
                if 'G' in cf_range and any(c.isdigit() for c in cf_range):
                    has_col_g_cf = True
            if has_col_a_cf and has_col_g_cf:
                cf_ok_count += 1

        if cf_ok_count >= 11:
            print(f"PASS: Component 6 — {cf_ok_count}/12 sheets have CF on weekend columns (0.10 pts)")
            total_score += 0.10
        elif cf_ok_count >= 6:
            partial = round(0.10 * cf_ok_count / 12, 3)
            print(f"PARTIAL: Component 6 — {cf_ok_count}/12 sheets ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 6 — Only {cf_ok_count}/12 sheets have CF on weekend columns")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: Navigation hyperlinks between months (0.10 pts)
    try:
        nav_ok_count = 0
        for i, month in enumerate(MONTH_NAMES):
            if month not in sheet_names:
                continue
            ws = wb[month]
            has_nav = False
            # Search for hyperlinks in the sheet
            for row in ws.iter_rows():
                for cell in row:
                    if cell.hyperlink:
                        has_nav = True
                        break
                if has_nav:
                    break
            # Also check for text containing ">>" or "<<" or other month names as navigation
            if not has_nav:
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            val_lower = cell.value.lower()
                            for other_month in MONTH_NAMES:
                                if other_month != month and other_month.lower() in val_lower:
                                    has_nav = True
                                    break
                        if has_nav:
                            break
                    if has_nav:
                        break
            if has_nav:
                nav_ok_count += 1

        if nav_ok_count >= 11:
            print(f"PASS: Component 7 — {nav_ok_count}/12 sheets have navigation (0.10 pts)")
            total_score += 0.10
        elif nav_ok_count >= 6:
            partial = round(0.10 * nav_ok_count / 12, 3)
            print(f"PARTIAL: Component 7 — {nav_ok_count}/12 sheets ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 7 — Only {nav_ok_count}/12 sheets have navigation")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
