"""
Reward Script: Create a tax calculation worksheet with progressive tax brackets and formatted breakdown.
Task ID: calc_gpm_071
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Tax bracket calculations D9:D15 correct
  Component 2 (0.20): Cumulative tax E9:E15 correct
  Component 3 (0.20): Summary rows 16-18 (Total Federal Tax, Effective Rate, Marginal Rate)
  Component 4 (0.15): E16 formatting (bold, 12pt, double underline) + number formats
  Component 5 (0.15): Active bracket row 11 highlighted yellow
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_071'


def persist_app_state(domain):
    """Attempt to save any unsaved LibreOffice state via Ctrl+S."""
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        import time
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for %s" % domain)
    except Exception as e:
        print("PERSIST_WARN: save hook failed: %s" % e)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print("CRITICAL: Cannot load file %s: %s" % (file_path, e))
        print("REWARD: 0.0")
        return 0.0

    if 'TaxCalc' not in wb.sheetnames:
        print("CRITICAL: Sheet 'TaxCalc' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['TaxCalc']

    # ----------------------------------------------------------------
    # Component 1: Tax bracket calculations D9:D15 (0.30 points)
    # Expected: D9=1160, D10=4266, D11=7315, D12-D15=0
    # These are the progressive tax amounts for taxable income = 80400
    # Bracket 1: 10% of 11600 = 1160
    # Bracket 2: 12% of (47150-11600) = 12% of 35550 = 4266
    # Bracket 3: 22% of (80400-47150) = 22% of 33250 = 7315
    # Brackets 4-7: 0 (taxable income doesn't reach these)
    # ----------------------------------------------------------------
    try:
        expected_d = {
            9: 1160, 10: 4266, 11: 7315,
            12: 0, 13: 0, 14: 0, 15: 0
        }
        d_pass_count = 0
        for row, expected_val in expected_d.items():
            actual = ws.cell(row=row, column=4).value  # column D
            if actual is not None:
                try:
                    if abs(float(actual) - expected_val) <= 1.0:
                        d_pass_count += 1
                    else:
                        print("FAIL: Component 1 — D%d expected %s, found %s" % (row, expected_val, actual))
                except (ValueError, TypeError):
                    print("FAIL: Component 1 — D%d not numeric: %s" % (row, actual))
            else:
                print("FAIL: Component 1 — D%d is None" % row)

        if d_pass_count == 7:
            print("PASS: Component 1 — All tax bracket calculations correct (0.30 pts)")
            total_score += 0.30
        elif d_pass_count >= 4:
            partial = 0.15
            print("PARTIAL: Component 1 — %d/7 bracket calcs correct (%.2f pts)" % (d_pass_count, partial))
            total_score += partial
        elif d_pass_count >= 1:
            partial = 0.05
            print("PARTIAL: Component 1 — %d/7 bracket calcs correct (%.2f pts)" % (d_pass_count, partial))
            total_score += partial
        else:
            print("FAIL: Component 1 — No bracket calculations found")
    except Exception as e:
        print("ERROR: Component 1 — %s" % e)

    # ----------------------------------------------------------------
    # Component 2: Cumulative tax E9:E15 (0.20 points)
    # Expected: E9=1160, E10=5426, E11=12741, E12-E15=12741
    # ----------------------------------------------------------------
    try:
        expected_e = {
            9: 1160, 10: 5426, 11: 12741,
            12: 12741, 13: 12741, 14: 12741, 15: 12741
        }
        e_pass_count = 0
        for row, expected_val in expected_e.items():
            actual = ws.cell(row=row, column=5).value  # column E
            if actual is not None:
                try:
                    if abs(float(actual) - expected_val) <= 1.0:
                        e_pass_count += 1
                    else:
                        print("FAIL: Component 2 — E%d expected %s, found %s" % (row, expected_val, actual))
                except (ValueError, TypeError):
                    print("FAIL: Component 2 — E%d not numeric: %s" % (row, actual))
            else:
                print("FAIL: Component 2 — E%d is None" % row)

        if e_pass_count == 7:
            print("PASS: Component 2 — All cumulative tax values correct (0.20 pts)")
            total_score += 0.20
        elif e_pass_count >= 4:
            partial = 0.10
            print("PARTIAL: Component 2 — %d/7 cumulative values correct (%.2f pts)" % (e_pass_count, partial))
            total_score += partial
        elif e_pass_count >= 1:
            partial = 0.05
            print("PARTIAL: Component 2 — %d/7 cumulative values correct (%.2f pts)" % (e_pass_count, partial))
            total_score += partial
        else:
            print("FAIL: Component 2 — No cumulative tax values found")
    except Exception as e:
        print("ERROR: Component 2 — %s" % e)

    # ----------------------------------------------------------------
    # Component 3: Summary rows 16-18 (0.20 points)
    # A16='Total Federal Tax', E16=12741
    # A17='Effective Rate', E17=formula '=E16/B4'
    # A18='Marginal Rate', E18=0.22
    # ----------------------------------------------------------------
    try:
        c3_points = 0.0

        # Check A16 label and E16 value
        a16_val = ws['A16'].value
        e16_val = ws['E16'].value
        if a16_val and 'total' in str(a16_val).lower() and 'tax' in str(a16_val).lower():
            if e16_val is not None:
                try:
                    if abs(float(e16_val) - 12741) <= 5.0:
                        c3_points += 0.08
                        print("PASS: Component 3a — Total Federal Tax label + value correct")
                    else:
                        print("FAIL: Component 3a — E16 expected ~12741, found %s" % e16_val)
                except (ValueError, TypeError):
                    # Could be a formula string
                    print("FAIL: Component 3a — E16 not numeric: %s" % repr(e16_val))
            else:
                print("FAIL: Component 3a — E16 is None")
        else:
            print("FAIL: Component 3a — A16 expected 'Total Federal Tax', found %s" % repr(a16_val))

        # Check A17 label and E17 formula
        a17_val = ws['A17'].value
        e17_val = ws['E17'].value
        if a17_val and 'effective' in str(a17_val).lower() and 'rate' in str(a17_val).lower():
            if e17_val is not None:
                e17_str = str(e17_val).upper().replace(' ', '')
                # Accept formula or computed value
                if '=E16/B4' in e17_str or '=E16/' in e17_str:
                    c3_points += 0.06
                    print("PASS: Component 3b — Effective Rate formula found: %s" % repr(e17_val))
                else:
                    # Maybe it's a computed value (should be ~0.1341)
                    try:
                        if abs(float(e17_val) - 0.1341) <= 0.01:
                            c3_points += 0.06
                            print("PASS: Component 3b — Effective Rate value correct: %s" % e17_val)
                        else:
                            print("FAIL: Component 3b — E17 value unexpected: %s" % repr(e17_val))
                    except (ValueError, TypeError):
                        print("FAIL: Component 3b — E17 unexpected: %s" % repr(e17_val))
            else:
                print("FAIL: Component 3b — E17 is None")
        else:
            print("FAIL: Component 3b — A17 expected 'Effective Rate', found %s" % repr(a17_val))

        # Check A18 label and E18 value
        a18_val = ws['A18'].value
        e18_val = ws['E18'].value
        if a18_val and 'marginal' in str(a18_val).lower() and 'rate' in str(a18_val).lower():
            if e18_val is not None:
                try:
                    if abs(float(e18_val) - 0.22) <= 0.01:
                        c3_points += 0.06
                        print("PASS: Component 3c — Marginal Rate value correct: %s" % e18_val)
                    else:
                        print("FAIL: Component 3c — E18 expected 0.22, found %s" % e18_val)
                except (ValueError, TypeError):
                    print("FAIL: Component 3c — E18 not numeric: %s" % repr(e18_val))
            else:
                print("FAIL: Component 3c — E18 is None")
        else:
            print("FAIL: Component 3c — A18 expected 'Marginal Rate', found %s" % repr(a18_val))

        if c3_points > 0:
            print("Component 3 total: %.2f pts" % c3_points)
            total_score += c3_points
        else:
            print("FAIL: Component 3 — No summary row checks passed")
    except Exception as e:
        print("ERROR: Component 3 — %s" % e)

    # ----------------------------------------------------------------
    # Component 4: E16 formatting — bold, 12pt, double underline + number formats (0.15 points)
    # Also check D/E number format $#,##0.00 and E17 0.00%
    # ----------------------------------------------------------------
    try:
        c4_points = 0.0

        e16_cell = ws['E16']
        # Check bold
        if e16_cell.font.bold:
            c4_points += 0.03
            print("PASS: Component 4a — E16 is bold")
        else:
            print("FAIL: Component 4a — E16 not bold")

        # Check font size 12
        if e16_cell.font.size is not None and abs(float(e16_cell.font.size) - 12.0) <= 0.5:
            c4_points += 0.03
            print("PASS: Component 4b — E16 font size is 12pt")
        else:
            print("FAIL: Component 4b — E16 font size expected 12, found %s" % e16_cell.font.size)

        # Check double underline
        if e16_cell.font.underline == 'double':
            c4_points += 0.03
            print("PASS: Component 4c — E16 has double underline")
        else:
            print("FAIL: Component 4c — E16 underline expected 'double', found %s" % e16_cell.font.underline)

        # Check E16 number format contains $ and #
        e16_fmt = ws['E16'].number_format
        if '$' in str(e16_fmt):
            c4_points += 0.03
            print("PASS: Component 4d — E16 number format is currency: %s" % e16_fmt)
        else:
            print("FAIL: Component 4d — E16 number format expected currency, found: %s" % e16_fmt)

        # Check E17 number format is percentage
        e17_fmt = ws['E17'].number_format
        if '%' in str(e17_fmt):
            c4_points += 0.03
            print("PASS: Component 4e — E17 number format is percentage: %s" % e17_fmt)
        else:
            print("FAIL: Component 4e — E17 number format expected percent, found: %s" % e17_fmt)

        if c4_points > 0:
            print("Component 4 total: %.2f pts" % c4_points)
            total_score += c4_points
        else:
            print("FAIL: Component 4 — No formatting checks passed")
    except Exception as e:
        print("ERROR: Component 4 — %s" % e)

    # ----------------------------------------------------------------
    # Component 5: Active bracket row 11 highlighted yellow (0.15 points)
    # Row 11 = bracket 3 (22%, $47,151-$100,525) is the active bracket
    # for taxable income of $80,400. Should have yellow fill (FFFFFF00).
    # ----------------------------------------------------------------
    try:
        c5_points = 0.0
        yellow_count = 0
        for col in range(1, 6):
            cell = ws.cell(row=11, column=col)
            fg = None
            try:
                fg = cell.fill.fgColor.rgb
            except Exception:
                pass
            if fg and 'FFFF00' in str(fg):
                yellow_count += 1

        if yellow_count >= 3:
            c5_points = 0.15
            print("PASS: Component 5 — Row 11 (active bracket) has yellow fill (%d/5 cols)" % yellow_count)
        elif yellow_count >= 1:
            c5_points = 0.07
            print("PARTIAL: Component 5 — Row 11 partially yellow (%d/5 cols)" % yellow_count)
        else:
            print("FAIL: Component 5 — Row 11 has no yellow fill")

        # Also verify that OTHER bracket rows (9,10,12-15) do NOT have yellow
        # (to confirm it's specifically the active bracket that's highlighted)
        other_yellow = 0
        for r in [9, 10, 12, 13, 14, 15]:
            cell = ws.cell(row=r, column=1)
            fg = None
            try:
                fg = cell.fill.fgColor.rgb
            except Exception:
                pass
            if fg and 'FFFF00' in str(fg):
                other_yellow += 1

        if other_yellow > 0 and c5_points > 0:
            # Non-active brackets also yellow — less meaningful highlight
            c5_points = max(c5_points - 0.05, 0.05)
            print("WARN: Component 5 — %d non-active bracket rows also have yellow fill" % other_yellow)

        if c5_points > 0:
            total_score += c5_points
            print("Component 5 total: %.2f pts" % c5_points)
    except Exception as e:
        print("ERROR: Component 5 — %s" % e)

    final_score = min(round(total_score, 2), 1.0)
    print("\nScore: %.2f/1.0" % total_score)
    print("REWARD: %.1f" % final_score)
    return final_score


# Entrypoint
persist_app_state("libreoffice_calc")

file_path = '%s/%s.xlsx' % (WORKDIR, TASK_ID)
if not os.path.exists(file_path):
    print("File not found: %s" % file_path)
    print("REWARD: 0.0")
else:
    verify_task(file_path)
