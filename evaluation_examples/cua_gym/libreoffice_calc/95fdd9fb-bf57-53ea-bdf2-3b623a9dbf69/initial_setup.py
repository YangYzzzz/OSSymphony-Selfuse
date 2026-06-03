"""
Initial Setup: Tax calculation worksheet - pre-task state with raw data only
Task ID: calc_gpm_071
Domain: libreoffice_calc

The initial state provides:
- Sheet 'TaxCalc' with a title header
- Input section with Filing Status, Gross Income, Deductions
- Tax bracket table with bracket definitions (Rate, Income Range)
- NO calculated tax amounts in D or E columns
- NO summary rows (Total Tax, Effective Rate, Marginal Rate)
- NO conditional formatting for active bracket or data bars
- The agent must: calculate taxes per bracket, add cumulative totals,
  add summary rows, and apply conditional formatting.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.worksheet.datavalidation import DataValidation

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_071'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'TaxCalc'

    # --- Colors ---
    dark_blue_fill = PatternFill(start_color="FF003366", end_color="FF003366", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    white_font = Font(color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # --- Title Row: Merged A1:F1 ---
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "Federal Income Tax Calculator - 2025 Filing"
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.fill = dark_blue_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Input Section (A3:B5) with yellow fill ---
    ws["A3"] = "Filing Status:"
    ws["A3"].fill = yellow_fill
    ws["B3"] = "Single"
    ws["B3"].fill = yellow_fill

    # Add dropdown validation for Filing Status
    dv = DataValidation(
        type="list",
        formula1='"Single,Married Filing Jointly,Head of Household"',
        allow_blank=False,
        showDropDown=False,
    )
    dv.add("B3")
    ws.add_data_validation(dv)

    ws["A4"] = "Gross Income:"
    ws["A4"].fill = yellow_fill
    ws["B4"] = 95000
    ws["B4"].fill = yellow_fill
    ws["B4"].number_format = '$#,##0'

    ws["A5"] = "Deductions:"
    ws["A5"].fill = yellow_fill
    ws["B5"] = 14600
    ws["B5"].fill = yellow_fill
    ws["B5"].number_format = '$#,##0'

    # --- B6: Taxable Income (formula) ---
    ws["A6"] = "Taxable Income"
    ws["A6"].font = Font(bold=True)
    ws["B6"] = "=B4-B5"
    ws["B6"].font = Font(bold=True)
    ws["B6"].number_format = '$#,##0'
    ws["B6"].border = thin_border

    # --- Tax Bracket Table Headers (A8:E8) ---
    headers = ['Bracket', 'Rate', 'Income Range', 'Tax in Bracket', 'Cumulative Tax']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = dark_blue_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # --- Tax Bracket Data (A9:C15) - 7 brackets ---
    brackets = [
        [1, 0.10, '$0 - $11,600'],
        [2, 0.12, '$11,601 - $47,150'],
        [3, 0.22, '$47,151 - $100,525'],
        [4, 0.24, '$100,526 - $191,950'],
        [5, 0.32, '$191,951 - $243,725'],
        [6, 0.35, '$243,726 - $609,350'],
        [7, 0.37, '$609,351+'],
    ]

    for i, (bracket_num, rate, income_range) in enumerate(brackets):
        row = 9 + i
        ws.cell(row=row, column=1, value=bracket_num)
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")

        ws.cell(row=row, column=2, value=rate)
        ws.cell(row=row, column=2).number_format = '0%'
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")

        ws.cell(row=row, column=3, value=income_range)
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")

        # D and E columns: empty (agent must fill these)
        ws.cell(row=row, column=4).border = thin_border
        ws.cell(row=row, column=4).number_format = '$#,##0.00'
        ws.cell(row=row, column=5).border = thin_border
        ws.cell(row=row, column=5).number_format = '$#,##0.00'

    # --- Column widths for readability ---
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 5

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
