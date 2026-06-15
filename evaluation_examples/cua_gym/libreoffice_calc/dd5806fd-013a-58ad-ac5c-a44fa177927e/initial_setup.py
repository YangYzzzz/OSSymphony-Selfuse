"""
Initial Setup: Split Timeline sheet view horizontally at row 11 and vertically at column D
Task ID: calc_sht_split_003
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_sht_split_003'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Timeline ---
    ws = wb.active
    ws.title = 'Timeline'

    # Column headers: ID, Name, Owner (A-C), then Jan-Dec (D-O)
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    headers = ['ID', 'Name', 'Owner'] + months
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFFFF')

    # --- Rows 2-10: First half of year projects (Jan-Jun focused) ---
    projects_h1 = [
        ('P-001', 'Website Redesign',       'Alice Morgan'),
        ('P-002', 'Mobile App Launch',      'Brian Torres'),
        ('P-003', 'Data Platform Upgrade',  'Cynthia Park'),
        ('P-004', 'CRM Integration',        'David Nguyen'),
        ('P-005', 'Security Audit',         'Elena Russo'),
        ('P-006', 'HR Portal Development',  'Frank Osei'),
        ('P-007', 'BI Dashboard v2',        'Grace Kim'),
        ('P-008', 'Supply Chain Tracker',   'Henry Zhao'),
        ('P-009', 'Customer Portal',        'Irene Cabral'),
    ]
    # Milestone marker data: 'X' means milestone present, '' means none
    milestones_h1 = [
        # Jan Feb Mar Apr May Jun Jul Aug Sep Oct Nov Dec
        ['X', 'X', '',  '',  '',  '',  '',  '',  '',  '',  '',  ''],   # P-001
        ['',  'X', 'X', '',  '',  '',  '',  '',  '',  '',  '',  ''],   # P-002
        ['',  '',  'X', 'X', '',  '',  '',  '',  '',  '',  '',  ''],   # P-003
        ['X', '',  '',  'X', 'X', '',  '',  '',  '',  '',  '',  ''],   # P-004
        ['',  'X', '',  '',  'X', 'X', '',  '',  '',  '',  '',  ''],   # P-005
        ['',  '',  'X', '',  '',  'X', '',  '',  '',  '',  '',  ''],   # P-006
        ['X', '',  '',  '',  '',  '',  'X', '',  '',  '',  '',  ''],   # P-007
        ['',  'X', '',  '',  '',  '',  '',  'X', '',  '',  '',  ''],   # P-008
        ['',  '',  '',  'X', '',  '',  '',  '',  'X', '',  '',  ''],   # P-009
    ]
    for r, (proj, milestones) in enumerate(zip(projects_h1, milestones_h1), 2):
        ws.cell(row=r, column=1, value=proj[0])
        ws.cell(row=r, column=2, value=proj[1])
        ws.cell(row=r, column=3, value=proj[2])
        for c, marker in enumerate(milestones, 4):
            ws.cell(row=r, column=c, value=marker)

    # --- Rows 11-20: Second half of year projects (Jul-Dec focused) ---
    projects_h2 = [
        ('P-010', 'ERP Migration Phase 2',   'James Okafor'),
        ('P-011', 'Cloud Cost Optimization', 'Karen Liu'),
        ('P-012', 'Compliance Framework',    'Liam Singh'),
        ('P-013', 'Marketing Automation',    'Maya Patel'),
        ('P-014', 'AI Analytics Pilot',      'Nathan Brooks'),
        ('P-015', 'Infrastructure Refresh',  'Olivia Jensen'),
        ('P-016', 'DevOps Pipeline Update',  'Paul Yamamoto'),
        ('P-017', 'Product Catalog Rewrite', 'Quinn Andersen'),
        ('P-018', 'Partner API Gateway',     'Rachel Ferreira'),
        ('P-019', 'Year-End Reporting Tool', 'Samuel Diallo'),
    ]
    milestones_h2 = [
        # Jan Feb Mar Apr May Jun Jul Aug Sep Oct Nov Dec
        ['',  '',  '',  '',  '',  '',  'X', 'X', '',  '',  '',  ''],   # P-010
        ['',  '',  '',  '',  '',  '',  '',  'X', 'X', '',  '',  ''],   # P-011
        ['',  '',  '',  '',  '',  '',  '',  '',  'X', 'X', '',  ''],   # P-012
        ['',  '',  '',  '',  '',  '',  'X', '',  '',  'X', 'X', ''],   # P-013
        ['',  '',  '',  '',  '',  '',  '',  'X', '',  '',  'X', 'X'],  # P-014
        ['',  '',  '',  '',  '',  '',  'X', '',  '',  '',  '',  'X'],  # P-015
        ['X', '',  '',  '',  '',  '',  '',  '',  'X', '',  '',  'X'],  # P-016
        ['',  '',  'X', '',  '',  '',  '',  '',  '',  'X', '',  'X'],  # P-017
        ['',  '',  '',  '',  'X', '',  '',  '',  '',  '',  'X', 'X'],  # P-018
        ['',  '',  '',  '',  '',  '',  '',  '',  '',  '',  '',  'X'],  # P-019
    ]
    for r, (proj, milestones) in enumerate(zip(projects_h2, milestones_h2), 11):
        ws.cell(row=r, column=1, value=proj[0])
        ws.cell(row=r, column=2, value=proj[1])
        ws.cell(row=r, column=3, value=proj[2])
        for c, marker in enumerate(milestones, 4):
            ws.cell(row=r, column=c, value=marker)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 20
    for col_letter in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
        ws.column_dimensions[col_letter].width = 5

    # IMPORTANT: No split/freeze applied — this is the initial state

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
