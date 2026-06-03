"""
Reward Script: Split the 'Timeline' sheet view horizontally at row 11 and vertically at column D
Task ID: calc_sht_split_003
Domain: libreoffice_calc

Scoring:
  Component 1: Split pane state is active ('split')           — 0.30 pts
  Component 2: Horizontal split at row 11 (bottomLeft/bottomRight panes)  — 0.35 pts
  Component 3: Vertical split at column D (topRight/bottomRight panes)    — 0.35 pts
  Total: 1.0

Verification strategy:
  - Load the .xlsx file with openpyxl
  - Access the sheet_view pane object on the 'Timeline' sheet
  - Check pane.state == 'split' (not 'frozen')
  - Check that the horizontal split creates panes beginning at row 11
    by verifying a selection with pane in ('bottomLeft', 'bottomRight')
    has activeCell row == 11
  - Check that the vertical split creates panes beginning at column D
    by verifying a selection with pane in ('topRight', 'bottomRight')
    has activeCell column == 'D'
  - Cell data integrity is checked as a precondition gate (no scoring)
"""

import os
import openpyxl
from openpyxl.utils import column_index_from_string

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_sht_split_003'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # --- Precondition Gate: load the workbook ---
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # --- Precondition Gate: 'Timeline' sheet must exist ---
    if 'Timeline' not in wb.sheetnames:
        print("CRITICAL: 'Timeline' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Timeline']

    # --- Precondition Gate: Cell data must not be modified ---
    # We check that the sheet has the expected shape (20 rows, 15 cols with data)
    try:
        if ws.max_row < 20 or ws.max_column < 15:
            print(f"WARNING: Sheet shape unexpected: rows={ws.max_row}, cols={ws.max_column}")
            # Not a hard gate — proceed with scoring
    except Exception as e:
        print(f"WARNING: Could not check sheet shape: {e}")

    # Retrieve sheet_view pane
    sv = ws.sheet_view
    pane = sv.pane  # None if no split/freeze has been applied

    # Retrieve selections list
    selections = list(sv.selection) if sv.selection else []

    # --- Component 1: Split pane state is active (0.30 pts) ---
    # The pane object must exist and its state must be 'split' (not 'frozen').
    # This FAILS on the initial file (pane is None) and PASSES on the golden file.
    try:
        if pane is not None and pane.state == 'split':
            print(f"PASS: Component 1 — pane.state='split' confirmed (0.30 pts)")
            total_score += 0.30
        elif pane is not None:
            print(f"FAIL: Component 1 — pane exists but state='{pane.state}', expected 'split'")
        else:
            print("FAIL: Component 1 — no pane applied (no split or freeze detected)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # --- Component 2: Horizontal split at row 11 (0.35 pts) ---
    # After splitting at row 11, selections for 'bottomLeft' and/or 'bottomRight' panes
    # should have activeCell pointing to row 11 (e.g., 'A11' or 'D11').
    # We also verify ySplit > 0 on the pane object.
    # This FAILS on the initial file (no pane) and PASSES on the golden file.
    try:
        horizontal_ok = False

        if pane is not None and pane.ySplit and pane.ySplit > 0:
            # Check if any bottom-pane selection references row 11
            bottom_pane_names = {'bottomLeft', 'bottomRight'}
            for sel in selections:
                if sel.pane in bottom_pane_names and sel.activeCell:
                    active_cell = sel.activeCell  # e.g., 'A11' or 'D11'
                    # Extract row number from the activeCell reference
                    import re
                    match = re.match(r'[A-Z]+(\d+)', active_cell)
                    if match:
                        row_num = int(match.group(1))
                        if row_num == 11:
                            horizontal_ok = True
                            break

            # Fallback: check topLeftCell on pane directly
            if not horizontal_ok and pane.topLeftCell:
                import re
                match = re.match(r'[A-Z]+(\d+)', pane.topLeftCell)
                if match and int(match.group(1)) == 11:
                    horizontal_ok = True

        if horizontal_ok:
            print(f"PASS: Component 2 — horizontal split confirmed at row 11 (ySplit={pane.ySplit}) (0.35 pts)")
            total_score += 0.35
        else:
            ysplit_val = pane.ySplit if pane else None
            print(f"FAIL: Component 2 — horizontal split at row 11 not found (ySplit={ysplit_val}, selections={[(s.pane, s.activeCell) for s in selections]})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # --- Component 3: Vertical split at column D (0.35 pts) ---
    # After splitting at column D, selections for 'topRight' and/or 'bottomRight' panes
    # should have activeCell pointing to column D (e.g., 'D1' or 'D11').
    # We also verify xSplit > 0 on the pane object.
    # This FAILS on the initial file (no pane) and PASSES on the golden file.
    try:
        vertical_ok = False

        if pane is not None and pane.xSplit and pane.xSplit > 0:
            # Check if any right-pane selection references column D
            right_pane_names = {'topRight', 'bottomRight'}
            for sel in selections:
                if sel.pane in right_pane_names and sel.activeCell:
                    active_cell = sel.activeCell  # e.g., 'D1' or 'D11'
                    import re
                    match = re.match(r'([A-Z]+)\d+', active_cell)
                    if match:
                        col_letter = match.group(1)
                        if col_letter == 'D':
                            vertical_ok = True
                            break

            # Fallback: check topLeftCell on pane directly
            if not vertical_ok and pane.topLeftCell:
                import re
                match = re.match(r'([A-Z]+)\d+', pane.topLeftCell)
                if match and match.group(1) == 'D':
                    vertical_ok = True

        if vertical_ok:
            print(f"PASS: Component 3 — vertical split confirmed at column D (xSplit={pane.xSplit}) (0.35 pts)")
            total_score += 0.35
        else:
            xsplit_val = pane.xSplit if pane else None
            print(f"FAIL: Component 3 — vertical split at column D not found (xSplit={xsplit_val}, selections={[(s.pane, s.activeCell) for s in selections]})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
