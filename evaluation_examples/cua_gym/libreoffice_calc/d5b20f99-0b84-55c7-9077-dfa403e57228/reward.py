"""
Reward Script: Set the height of row 1 to 30 points
Task ID: calc_fmt_row_height_specific_048
Domain: libreoffice_calc
Scoring:
  Component 1 (0.7): Row 1 height is exactly 30 points (the primary task change)
  Component 2 (0.3): Row 1 height is 30 pts AND all other rows unchanged AND
                     cell contents are intact (compound check anchoring the change)
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmt_row_height_specific_048'

# Expected header values (task context: row 1 is the header row)
EXPECTED_HEADERS = {1: 'Date', 2: 'Activity', 3: 'Hours', 4: 'Notes'}


def check_row1_height(ws):
    """Return True if row 1 height is set to approximately 30 points."""
    h = ws.row_dimensions[1].height
    return h is not None and abs(float(h) - 30.0) <= 0.5


def check_other_rows_unchanged(ws):
    """Return (True, []) if rows 2-20 all have default height, else (False, changed_list)."""
    changed = []
    for row_idx in range(2, 21):
        h = ws.row_dimensions[row_idx].height
        if h is not None:
            changed.append((row_idx, h))
    return len(changed) == 0, changed


def check_headers_intact(ws):
    """Return (True, []) if row 1 header values match expected, else (False, diffs)."""
    diffs = []
    for col_idx, expected_val in EXPECTED_HEADERS.items():
        actual_val = ws.cell(row=1, column=col_idx).value
        if actual_val != expected_val:
            diffs.append((col_idx, expected_val, actual_val))
    return len(diffs) == 0, diffs


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Task: Set the height of row 1 to 30 points in the 'Weekly Report' sheet.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: sheet 'Weekly Report' must exist
    if 'Weekly Report' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Weekly Report' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Weekly Report']

    # Component 1: Row 1 height is set to 30 points (0.7 points)
    # In the initial file, row 1 height is None (default/auto).
    # After the task, row 1 height should be 30.0 pt.
    # This FAILS on initial (height=None) and PASSES on golden (height=30.0).
    try:
        row1_h = ws.row_dimensions[1].height
        if check_row1_height(ws):
            print(f"PASS: Component 1 — Row 1 height is {row1_h} pts (expected ~30.0 pts) (0.7 pts)")
            total_score += 0.7
        else:
            print(f"FAIL: Component 1 — Row 1 height is {row1_h!r}, expected 30.0 pts")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not check row 1 height: {e}")

    # Component 2: Row 1 height is 30 pts AND no other rows changed AND headers intact (0.3 points)
    # This is a compound check anchored to the task change. The primary condition
    # (row 1 = 30 pt) FAILS on initial file, making this compound check also fail on initial.
    # It PASSES on golden only if row 1 was correctly set while preserving data integrity.
    try:
        row1_ok = check_row1_height(ws)
        other_rows_ok, changed_rows = check_other_rows_unchanged(ws)
        headers_ok, header_diffs = check_headers_intact(ws)

        if row1_ok and other_rows_ok and headers_ok:
            print(f"PASS: Component 2 — Row 1=30 pts, rows 2-20 unchanged, headers intact (0.3 pts)")
            total_score += 0.3
        elif not row1_ok:
            print(f"FAIL: Component 2 — Row 1 height not at 30 pts (anchor condition failed)")
        elif not other_rows_ok:
            print(f"FAIL: Component 2 — Other row heights were unexpectedly modified: {changed_rows}")
        elif not headers_ok:
            print(f"FAIL: Component 2 — Header cell contents were modified: {header_diffs}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
