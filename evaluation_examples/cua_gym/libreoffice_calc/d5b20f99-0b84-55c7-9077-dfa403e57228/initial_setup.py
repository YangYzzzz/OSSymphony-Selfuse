"""
Initial Setup: Set row 1 height in a Weekly Report spreadsheet
Task ID: calc_fmt_row_height_specific_048
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_row_height_specific_048'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Weekly Report ---
    ws = wb.active
    ws.title = 'Weekly Report'

    # Headers in row 1 (no special formatting, default height)
    headers = ['Date', 'Activity', 'Hours', 'Notes']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Data rows 2-20 (realistic content, default row heights)
    activities = [
        'Project kickoff meeting',
        'Backend API development',
        'Code review session',
        'Client presentation prep',
        'Database schema design',
        'Frontend UI implementation',
        'Sprint planning meeting',
        'Bug fixing and testing',
        'Documentation writing',
        'Stakeholder sync call',
        'Performance optimization',
        'Team retrospective',
        'Requirements analysis',
        'Deployment pipeline setup',
        'Security audit review',
        'Data migration scripts',
        'Unit test writing',
        'Design system updates',
        'Architecture discussion',
    ]

    notes_list = [
        'All team members attended',
        'Completed user authentication module',
        'Reviewed 3 pull requests',
        'Slides finalized, feedback incorporated',
        'Normalized to 3NF',
        'Dashboard components completed',
        'Story points estimated for sprint 7',
        'Fixed 5 critical bugs',
        'API docs updated on Confluence',
        'Discussed Q3 roadmap changes',
        'Reduced load time by 40%',
        'Action items logged in Jira',
        'Gathered requirements from 4 stakeholders',
        'CI/CD pipeline with GitHub Actions',
        'No critical vulnerabilities found',
        'Migrated 50k records successfully',
        'Coverage increased to 82%',
        'Updated color palette and typography',
        'Decided on microservices approach',
    ]

    start_date = date(2025, 3, 3)  # Monday
    row = 2
    for i, (activity, note) in enumerate(zip(activities, notes_list)):
        # Compute date: Mon-Fri of each week, skip weekends
        day_offset = (i // 5) * 7 + (i % 5)
        current_date = start_date + timedelta(days=day_offset)
        hours = [2.5, 3.0, 1.5, 4.0, 2.0, 3.5, 1.0, 5.0, 2.5, 1.5,
                 3.0, 2.0, 4.5, 3.5, 2.0, 5.0, 3.0, 1.5, 4.0][i]
        ws.cell(row=row, column=1, value=current_date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value=activity)
        ws.cell(row=row, column=3, value=hours)
        ws.cell(row=row, column=4, value=note)
        row += 1

    # Column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 38

    # NOTE: Row 1 height is intentionally left at default (NOT set to 30pt)
    # The task requires the agent to set it to 30pt

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
