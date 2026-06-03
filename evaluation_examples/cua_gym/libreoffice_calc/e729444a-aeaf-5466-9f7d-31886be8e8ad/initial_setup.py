"""
Initial Setup: Add a data validation dropdown list to cells D2:D50 in the 'Tasks' sheet.
Task ID: calc_dop_validate_dropdown_019
Domain: libreoffice_calc

Creates a Tasks sheet with 49 task records. Column D contains inconsistent free-form
status text (not yet validated). No data validation is applied initially.
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_dop_validate_dropdown_019'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Tasks ---
    ws = wb.active
    ws.title = 'Tasks'

    # Headers in row 1
    headers = ['Task ID', 'Task Name', 'Assignee', 'Status', 'Priority', 'Due Date']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # 49 task records with realistic content
    # Column D has inconsistent free-form text (no validation yet)
    task_data = [
        ('T-001', 'Redesign landing page',       'Emma Larson',     'done',         'High',   '2025-04-10'),
        ('T-002', 'Update user authentication',  'Carlos Rivera',   'in progress',  'High',   '2025-04-15'),
        ('T-003', 'Fix payment gateway bug',     'Priya Nair',      'not started',  'High',   '2025-04-08'),
        ('T-004', 'Write API documentation',     'Tom Hadley',      'started',      'Medium', '2025-04-20'),
        ('T-005', 'Set up CI/CD pipeline',       'Emma Larson',     'blocked',      'High',   '2025-04-12'),
        ('T-006', 'Migrate database to AWS',     'Carlos Rivera',   'not yet',      'High',   '2025-04-18'),
        ('T-007', 'Conduct UX research',         'Aisha Okonkwo',   'done',         'Medium', '2025-03-28'),
        ('T-008', 'Implement dark mode',         'Tom Hadley',      'in progress',  'Low',    '2025-05-01'),
        ('T-009', 'Optimize search indexing',    'Priya Nair',      'todo',         'Medium', '2025-04-22'),
        ('T-010', 'Create onboarding flow',      'Marcus Bell',     'started',      'High',   '2025-04-16'),
        ('T-011', 'Audit accessibility issues',  'Aisha Okonkwo',   'not yet',      'Medium', '2025-04-30'),
        ('T-012', 'Deploy hotfix v2.1.3',        'Emma Larson',     'done',         'High',   '2025-03-30'),
        ('T-013', 'Build admin dashboard',       'Carlos Rivera',   'in progress',  'High',   '2025-05-05'),
        ('T-014', 'Revamp email templates',      'Sofia Marchetti', 'not started',  'Low',    '2025-05-10'),
        ('T-015', 'Integrate Stripe v3',         'Tom Hadley',      'blocked',      'High',   '2025-04-14'),
        ('T-016', 'Add multi-language support',  'Marcus Bell',     'todo',         'Medium', '2025-05-20'),
        ('T-017', 'Fix iOS rendering bug',       'Priya Nair',      'done',         'High',   '2025-04-02'),
        ('T-018', 'Run load testing',            'Aisha Okonkwo',   'in progress',  'High',   '2025-04-17'),
        ('T-019', 'Update privacy policy',       'Sofia Marchetti', 'started',      'Low',    '2025-04-25'),
        ('T-020', 'Refactor cart module',        'Emma Larson',     'not yet',      'Medium', '2025-05-03'),
        ('T-021', 'Set up error monitoring',     'Carlos Rivera',   'done',         'High',   '2025-04-05'),
        ('T-022', 'Design icon set',             'Sofia Marchetti', 'in progress',  'Medium', '2025-04-28'),
        ('T-023', 'Write unit tests for auth',   'Tom Hadley',      'todo',         'High',   '2025-04-19'),
        ('T-024', 'Implement rate limiting',     'Priya Nair',      'not started',  'High',   '2025-04-21'),
        ('T-025', 'Create analytics reports',    'Marcus Bell',     'started',      'Medium', '2025-05-07'),
        ('T-026', 'Fix CSV export bug',          'Aisha Okonkwo',   'done',         'Medium', '2025-04-03'),
        ('T-027', 'Add social login',            'Emma Larson',     'in progress',  'Medium', '2025-05-12'),
        ('T-028', 'Upgrade Node.js version',     'Carlos Rivera',   'blocked',      'High',   '2025-04-11'),
        ('T-029', 'Improve mobile checkout UX',  'Sofia Marchetti', 'not yet',      'High',   '2025-05-02'),
        ('T-030', 'Clean up dead code',          'Tom Hadley',      'todo',         'Low',    '2025-05-15'),
        ('T-031', 'Add 2FA support',             'Priya Nair',      'done',         'High',   '2025-04-06'),
        ('T-032', 'Localize billing module',     'Marcus Bell',     'started',      'Medium', '2025-05-18'),
        ('T-033', 'Performance audit Q2',        'Aisha Okonkwo',   'in progress',  'High',   '2025-04-23'),
        ('T-034', 'Create help center articles', 'Sofia Marchetti', 'not started',  'Low',    '2025-05-22'),
        ('T-035', 'Fix cookie consent popup',    'Emma Larson',     'done',         'Medium', '2025-04-07'),
        ('T-036', 'Set up staging environment',  'Carlos Rivera',   'todo',         'High',   '2025-04-26'),
        ('T-037', 'Rewrite logging system',      'Tom Hadley',      'not yet',      'Medium', '2025-05-08'),
        ('T-038', 'Implement PDF export',        'Priya Nair',      'in progress',  'Medium', '2025-05-01'),
        ('T-039', 'Update dependencies',         'Marcus Bell',     'done',         'Low',    '2025-04-04'),
        ('T-040', 'Add push notifications',      'Aisha Okonkwo',   'started',      'Medium', '2025-05-14'),
        ('T-041', 'Create data backup job',      'Sofia Marchetti', 'blocked',      'High',   '2025-04-13'),
        ('T-042', 'Design new pricing page',     'Emma Larson',     'not started',  'High',   '2025-05-06'),
        ('T-043', 'Integrate Google Analytics',  'Carlos Rivera',   'done',         'Medium', '2025-04-01'),
        ('T-044', 'Fix broken pagination',       'Tom Hadley',      'todo',         'Medium', '2025-04-27'),
        ('T-045', 'Refactor database schema',    'Priya Nair',      'in progress',  'High',   '2025-05-09'),
        ('T-046', 'Add webhook support',         'Marcus Bell',     'not yet',      'Medium', '2025-05-16'),
        ('T-047', 'Review GDPR compliance',      'Aisha Okonkwo',   'done',         'High',   '2025-04-09'),
        ('T-048', 'Rebuild search UI',           'Sofia Marchetti', 'started',      'Medium', '2025-05-19'),
        ('T-049', 'Conduct security pen test',   'Emma Larson',     'blocked',      'High',   '2025-04-24'),
    ]

    for r, row_data in enumerate(task_data, 2):
        ws.cell(row=r, column=1, value=row_data[0])   # Task ID
        ws.cell(row=r, column=2, value=row_data[1])   # Task Name
        ws.cell(row=r, column=3, value=row_data[2])   # Assignee
        ws.cell(row=r, column=4, value=row_data[3])   # Status (free-form, no validation)
        ws.cell(row=r, column=5, value=row_data[4])   # Priority
        ws.cell(row=r, column=6, value=row_data[5])   # Due Date

    # Set column widths for readability
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
