"""
Reward Script: SPC (Statistical Process Control) chart data table
Task ID: calc_ops_047
Domain: libreoffice_calc
Scoring:
  Component 1 — Mean formulas in C2:C21 (0.3 pts)
  Component 2 — UCL formulas in D2:D21 (0.2 pts)
  Component 3 — LCL formulas in E2:E21 (0.2 pts)
  Component 4 — Line chart with 4 series (0.3 pts)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_047'


def normalize_formula(f):
    """Normalize a formula string for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ""
    return f.upper().replace(" ", "")


def is_average_formula(f):
    """Check if formula computes AVERAGE($B$2:$B$21)."""
    nf = normalize_formula(f)
    # Accept =AVERAGE($B$2:$B$21) exactly
    return nf == "=AVERAGE($B$2:$B$21)"


def is_ucl_formula(f):
    """Check if formula computes Mean + 3*STDEV($B$2:$B$21).
    Accept various orderings like =AVERAGE(...)+3*STDEV(...) or =3*STDEV(...)+AVERAGE(...)
    """
    nf = normalize_formula(f)
    # Must start with =
    if not nf.startswith("="):
        return False
    expr = nf[1:]
    # Pattern: AVERAGE($B$2:$B$21)+3*STDEV($B$2:$B$21)
    # Also accept AVERAGE($B$2:$B$21)+STDEV($B$2:$B$21)*3
    patterns = [
        r"^AVERAGE\(\$B\$2:\$B\$21\)\+3\*STDEV\(\$B\$2:\$B\$21\)$",
        r"^AVERAGE\(\$B\$2:\$B\$21\)\+STDEV\(\$B\$2:\$B\$21\)\*3$",
        r"^3\*STDEV\(\$B\$2:\$B\$21\)\+AVERAGE\(\$B\$2:\$B\$21\)$",
        r"^STDEV\(\$B\$2:\$B\$21\)\*3\+AVERAGE\(\$B\$2:\$B\$21\)$",
    ]
    for pat in patterns:
        if re.match(pat, expr):
            return True
    return False


def is_lcl_formula(f):
    """Check if formula computes Mean - 3*STDEV($B$2:$B$21).
    Accept: AVERAGE($B$2:$B$21)-3*STDEV($B$2:$B$21)
    """
    nf = normalize_formula(f)
    if not nf.startswith("="):
        return False
    expr = nf[1:]
    patterns = [
        r"^AVERAGE\(\$B\$2:\$B\$21\)-3\*STDEV\(\$B\$2:\$B\$21\)$",
        r"^AVERAGE\(\$B\$2:\$B\$21\)-STDEV\(\$B\$2:\$B\$21\)\*3$",
    ]
    for pat in patterns:
        if re.match(pat, expr):
            return True
    return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet 'SPC' exists
    if 'SPC' not in wb.sheetnames:
        print("CRITICAL: Sheet 'SPC' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['SPC']

    # Component 1: Mean formulas in C2:C21 (0.3 points)
    # Each of the 20 cells should contain =AVERAGE($B$2:$B$21)
    try:
        mean_correct = 0
        for r in range(2, 22):
            val = ws.cell(row=r, column=3).value
            if is_average_formula(val):
                mean_correct += 1
            else:
                if mean_correct < 5:  # only print first few failures
                    print(f"  DETAIL: C{r} = {val!r} (not AVERAGE formula)")
        ratio = mean_correct / 20.0
        if ratio >= 0.9:
            pts = 0.3
        elif ratio >= 0.5:
            pts = round(0.3 * ratio, 2)
        else:
            pts = 0.0
        if pts > 0:
            print(f"PASS: Component 1 — Mean formulas: {mean_correct}/20 correct ({pts} pts)")
            total_score += pts
        else:
            print(f"FAIL: Component 1 — Mean formulas: {mean_correct}/20 correct")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: UCL formulas in D2:D21 (0.2 points)
    # Each should contain =AVERAGE($B$2:$B$21)+3*STDEV($B$2:$B$21)
    try:
        ucl_correct = 0
        for r in range(2, 22):
            val = ws.cell(row=r, column=4).value
            if is_ucl_formula(val):
                ucl_correct += 1
            else:
                if ucl_correct < 5:
                    print(f"  DETAIL: D{r} = {val!r} (not UCL formula)")
        ratio = ucl_correct / 20.0
        if ratio >= 0.9:
            pts = 0.2
        elif ratio >= 0.5:
            pts = round(0.2 * ratio, 2)
        else:
            pts = 0.0
        if pts > 0:
            print(f"PASS: Component 2 — UCL formulas: {ucl_correct}/20 correct ({pts} pts)")
            total_score += pts
        else:
            print(f"FAIL: Component 2 — UCL formulas: {ucl_correct}/20 correct")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: LCL formulas in E2:E21 (0.2 points)
    # Each should contain =AVERAGE($B$2:$B$21)-3*STDEV($B$2:$B$21)
    try:
        lcl_correct = 0
        for r in range(2, 22):
            val = ws.cell(row=r, column=5).value
            if is_lcl_formula(val):
                lcl_correct += 1
            else:
                if lcl_correct < 5:
                    print(f"  DETAIL: E{r} = {val!r} (not LCL formula)")
        ratio = lcl_correct / 20.0
        if ratio >= 0.9:
            pts = 0.2
        elif ratio >= 0.5:
            pts = round(0.2 * ratio, 2)
        else:
            pts = 0.0
        if pts > 0:
            print(f"PASS: Component 3 — LCL formulas: {lcl_correct}/20 correct ({pts} pts)")
            total_score += pts
        else:
            print(f"FAIL: Component 3 — LCL formulas: {lcl_correct}/20 correct")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Line chart with 4 data series (0.3 points)
    # The task requires a line chart showing measurement, mean, UCL, and LCL
    try:
        charts = ws._charts
        if len(charts) == 0:
            print("FAIL: Component 4 — No charts found on SPC sheet")
        else:
            # Count line charts and max series across all charts
            line_chart_count = sum(1 for ch in charts if isinstance(ch, openpyxl.chart.LineChart))
            max_series_on_line = max(
                (len(ch.series) for ch in charts if isinstance(ch, openpyxl.chart.LineChart)),
                default=0
            )
            max_series_any = max(len(ch.series) for ch in charts)
            for ch in charts:
                print(f"  DETAIL: Chart type={ch.__class__.__name__}, series={len(ch.series)}")

            if line_chart_count > 0 and max_series_on_line >= 4:
                print(f"PASS: Component 4 — Line chart with >= 4 series ({0.3} pts)")
                total_score += 0.3
            elif line_chart_count > 0:
                # Partial credit: line chart exists but not all 4 series
                print(f"PARTIAL: Component 4 — Line chart found but < 4 series ({0.15} pts)")
                total_score += 0.15
            elif max_series_any >= 4:
                # Has a chart with 4 series but not LineChart type
                print(f"PARTIAL: Component 4 — Chart with 4 series found but not LineChart ({0.15} pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 4 — No suitable chart found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved LibreOffice state before verification
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
