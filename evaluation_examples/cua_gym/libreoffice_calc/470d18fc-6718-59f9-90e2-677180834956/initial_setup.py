"""
Initial Setup: SPC (Statistical Process Control) chart data table
Task ID: calc_ops_047
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_047'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'SPC'

    # Headers
    headers = ['Sample', 'Measurement', 'Mean', 'UCL', 'LCL']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Measurement data as specified
    measurements = [
        10.2, 9.8, 10.1, 10.5, 9.7,
        10.3, 10.0, 9.9, 10.4, 10.1,
        9.6, 10.2, 10.3, 9.8, 10.0,
        10.5, 9.7, 10.1, 10.4, 9.9,
    ]

    for i in range(20):
        row = i + 2
        ws.cell(row=row, column=1, value=f'Sample {i + 1}')
        ws.cell(row=row, column=2, value=measurements[i])
        # Columns C, D, E intentionally left empty - task is to add formulas

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Open in LibreOffice Calc for GUI-ready state
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
