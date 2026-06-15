"""
Initial Setup: Capacity Planning Model
Task ID: calc_ops_production_capacity_027
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_production_capacity_027'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: WorkCenters ---
    ws1 = wb.active
    ws1.title = 'WorkCenters'

    # Headers: Work Center (A), Machines (B), Hours per Machine per Week (C),
    #          Gross Capacity (D - empty), Planned Maintenance Hrs (E), Net Capacity (F - empty)
    headers = ['Work Center', 'Machines', 'Hours per Machine per Week',
               'Gross Capacity', 'Planned Maintenance Hrs', 'Net Capacity']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # 8 work centers: A-C and E filled, D and F left EMPTY (task requires computing these)
    # Data: Work Center, Machines, Hours/Machine/Week, (D=empty), Planned Maintenance, (F=empty)
    work_centers_data = [
        ('Stamping Press',      4, 80, 12.0),
        ('Welding Station',     6, 80, 16.0),
        ('CNC Machining',       5, 80, 10.0),
        ('Assembly Line A',     8, 80, 20.0),
        ('Paint Booth',         2, 80,  8.0),
        ('Quality Inspection',  3, 80,  6.0),
        ('Packaging',           4, 80,  8.0),
        ('Final Assembly',      6, 80, 14.0),
    ]

    for r, (name, machines, hrs, maint) in enumerate(work_centers_data, 2):
        ws1.cell(row=r, column=1, value=name)
        ws1.cell(row=r, column=2, value=machines)
        ws1.cell(row=r, column=3, value=hrs)
        # Column D (Gross Capacity) — intentionally EMPTY
        ws1.cell(row=r, column=5, value=maint)
        # Column F (Net Capacity) — intentionally EMPTY

    # --- Sheet 2: DemandLoad ---
    ws2 = wb.create_sheet('DemandLoad')

    demand_headers = ['Work Center', 'Demand Hours This Week']
    for col, h in enumerate(demand_headers, 1):
        ws2.cell(row=1, column=col, value=h)

    # Demand hours — some will be over capacity, some under
    demand_data = [
        ('Stamping Press',     290),   # capacity = 4*80-12=308  → under
        ('Welding Station',    465),   # capacity = 6*80-16=464  → OVER
        ('CNC Machining',      390),   # capacity = 5*80-10=390  → exactly at (0 surplus)
        ('Assembly Line A',    625),   # capacity = 8*80-20=620  → OVER
        ('Paint Booth',        148),   # capacity = 2*80-8=152   → under
        ('Quality Inspection', 228),   # capacity = 3*80-6=234   → under
        ('Packaging',          310),   # capacity = 4*80-8=312   → under
        ('Final Assembly',     476),   # capacity = 6*80-14=466  → OVER
    ]

    for r, (name, demand) in enumerate(demand_data, 2):
        ws2.cell(row=r, column=1, value=name)
        ws2.cell(row=r, column=2, value=demand)

    # --- Sheet 3: CapacityAnalysis ---
    ws3 = wb.create_sheet('CapacityAnalysis')

    analysis_headers = ['Work Center', 'Net Capacity', 'Demand Hours',
                        'Surplus/Deficit', 'Utilization %', 'Status']
    for col, h in enumerate(analysis_headers, 1):
        ws3.cell(row=1, column=col, value=h)

    # Only work center names in column A; all other columns intentionally EMPTY
    wc_names = [row[0] for row in work_centers_data]
    for r, name in enumerate(wc_names, 2):
        ws3.cell(row=r, column=1, value=name)
        # Columns B-F intentionally EMPTY

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('WorkCenters sheet: A-C and E filled; D (Gross Capacity) and F (Net Capacity) empty')
    print('DemandLoad sheet: A-B filled with 8 work centers and demand hours')
    print('CapacityAnalysis sheet: A filled with work center names; B-F empty')


create_initial()
