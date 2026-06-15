"""
Reward Script: Capacity Planning Model for Production Work Centers
Task ID: calc_ops_production_capacity_027
Domain: libreoffice_calc

Scoring rubric:
  Component 1: WorkCenters Gross Capacity formulas (D2:D9 = =B*C)         — 0.25 pts
  Component 2: WorkCenters Net Capacity formulas (F2:F9 = =D-E)            — 0.25 pts
  Component 3: CapacityAnalysis formulas (B-F columns populated correctly) — 0.30 pts
  Component 4: Red conditional formatting for OVER CAPACITY rows            — 0.20 pts
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_production_capacity_027'


def check_formula_pattern(value, pattern_regex):
    """Check if a cell value matches a formula pattern (case-insensitive)."""
    if not isinstance(value, str):
        return False
    # Remove all whitespace for comparison
    normalized = value.replace(' ', '').upper()
    return bool(re.match(pattern_regex, normalized, re.IGNORECASE))


def verify_task(file_path):
    """
    Verify capacity planning model task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook — precondition gate
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify required sheets exist — precondition gate
    required_sheets = ['WorkCenters', 'DemandLoad', 'CapacityAnalysis']
    for sheet_name in required_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"CRITICAL: Required sheet '{sheet_name}' not found in workbook.")
            print("REWARD: 0.0")
            return 0.0

    ws_wc = wb['WorkCenters']
    ws_ca = wb['CapacityAnalysis']

    # -------------------------------------------------------------------------
    # Component 1: WorkCenters Gross Capacity formulas in D2:D9 (0.25 pts)
    # Each row D = =B*C (gross capacity = machines * hours per machine per week)
    # This checks the task-introduced change: D column was empty in initial file
    # -------------------------------------------------------------------------
    try:
        gross_pass = 0
        gross_total = 8
        for row in range(2, 10):
            d_val = ws_wc.cell(row=row, column=4).value  # Column D
            b_col = row
            # Accept formulas like =B2*C2 or =C2*B2 (case-insensitive, no spaces)
            if isinstance(d_val, str):
                normalized = d_val.replace(' ', '').upper()
                # Pattern: =B{row}*C{row} or =C{row}*B{row}
                if (normalized == f'=B{row}*C{row}' or
                        normalized == f'=C{row}*B{row}' or
                        f'B{row}*C{row}' in normalized or
                        f'C{row}*B{row}' in normalized):
                    gross_pass += 1
                else:
                    print(f"FAIL: WorkCenters D{row} has unexpected formula: {repr(d_val)}")
            else:
                print(f"FAIL: WorkCenters D{row} is not a formula string: {repr(d_val)}")

        if gross_pass == gross_total:
            print(f"PASS: Component 1 — All {gross_total} Gross Capacity formulas in D2:D9 are correct (0.25 pts)")
            total_score += 0.25
        elif gross_pass >= 4:
            partial = round(0.25 * (gross_pass / gross_total), 4)
            print(f"PARTIAL: Component 1 — {gross_pass}/{gross_total} Gross Capacity formulas correct (+{partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {gross_pass}/{gross_total} Gross Capacity formulas are correct")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: WorkCenters Net Capacity formulas in F2:F9 (0.25 pts)
    # Each row F = =D-E (net capacity = gross capacity - planned maintenance)
    # This checks the task-introduced change: F column was empty in initial file
    # -------------------------------------------------------------------------
    try:
        net_pass = 0
        net_total = 8
        for row in range(2, 10):
            f_val = ws_wc.cell(row=row, column=6).value  # Column F
            if isinstance(f_val, str):
                normalized = f_val.replace(' ', '').upper()
                # Pattern: =D{row}-E{row}
                if normalized == f'=D{row}-E{row}' or f'D{row}-E{row}' in normalized:
                    net_pass += 1
                else:
                    print(f"FAIL: WorkCenters F{row} has unexpected formula: {repr(f_val)}")
            else:
                print(f"FAIL: WorkCenters F{row} is not a formula string: {repr(f_val)}")

        if net_pass == net_total:
            print(f"PASS: Component 2 — All {net_total} Net Capacity formulas in F2:F9 are correct (0.25 pts)")
            total_score += 0.25
        elif net_pass >= 4:
            partial = round(0.25 * (net_pass / net_total), 4)
            print(f"PARTIAL: Component 2 — {net_pass}/{net_total} Net Capacity formulas correct (+{partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {net_pass}/{net_total} Net Capacity formulas are correct")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: CapacityAnalysis formulas in B2:F9 (0.30 pts)
    # B = VLOOKUP to WorkCenters!F for net capacity
    # C = VLOOKUP to DemandLoad!B for demand hours
    # D = =B-C (surplus/deficit)
    # E = =C/B (utilization %)
    # F = =IF(D<0,"OVER CAPACITY","OK")
    # All columns were empty in initial file
    # -------------------------------------------------------------------------
    try:
        ca_checks = {
            'B_vlookup': 0,  # Net Capacity VLOOKUP
            'C_vlookup': 0,  # Demand Hours VLOOKUP
            'D_formula': 0,  # Surplus/Deficit
            'E_formula': 0,  # Utilization %
            'F_formula': 0,  # Status IF
        }
        ca_total_per = 8

        for row in range(2, 10):
            b_val = ws_ca.cell(row=row, column=2).value
            c_val = ws_ca.cell(row=row, column=3).value
            d_val = ws_ca.cell(row=row, column=4).value
            e_val = ws_ca.cell(row=row, column=5).value
            f_val = ws_ca.cell(row=row, column=6).value

            # B: VLOOKUP referencing WorkCenters!F (6th column)
            if isinstance(b_val, str):
                b_norm = b_val.replace(' ', '').upper()
                # Should be VLOOKUP referencing WorkCenters and returning column 6
                if 'VLOOKUP' in b_norm and 'WORKCENTERS' in b_norm and '6' in b_norm:
                    ca_checks['B_vlookup'] += 1
                else:
                    print(f"FAIL: CapacityAnalysis B{row} unexpected formula: {repr(b_val)}")
            else:
                print(f"FAIL: CapacityAnalysis B{row} not a formula: {repr(b_val)}")

            # C: VLOOKUP referencing DemandLoad!B (2nd column)
            if isinstance(c_val, str):
                c_norm = c_val.replace(' ', '').upper()
                if 'VLOOKUP' in c_norm and 'DEMANDLOAD' in c_norm and '2' in c_norm:
                    ca_checks['C_vlookup'] += 1
                else:
                    print(f"FAIL: CapacityAnalysis C{row} unexpected formula: {repr(c_val)}")
            else:
                print(f"FAIL: CapacityAnalysis C{row} not a formula: {repr(c_val)}")

            # D: =B-C (surplus/deficit)
            if isinstance(d_val, str):
                d_norm = d_val.replace(' ', '').upper()
                if d_norm == f'=B{row}-C{row}' or f'B{row}-C{row}' in d_norm:
                    ca_checks['D_formula'] += 1
                else:
                    print(f"FAIL: CapacityAnalysis D{row} unexpected formula: {repr(d_val)}")
            else:
                print(f"FAIL: CapacityAnalysis D{row} not a formula: {repr(d_val)}")

            # E: =C/B (utilization %)
            if isinstance(e_val, str):
                e_norm = e_val.replace(' ', '').upper()
                if e_norm == f'=C{row}/B{row}' or f'C{row}/B{row}' in e_norm:
                    ca_checks['E_formula'] += 1
                else:
                    print(f"FAIL: CapacityAnalysis E{row} unexpected formula: {repr(e_val)}")
            else:
                print(f"FAIL: CapacityAnalysis E{row} not a formula: {repr(e_val)}")

            # F: =IF(D<0,"OVER CAPACITY","OK") or equivalent
            if isinstance(f_val, str):
                f_norm = f_val.replace(' ', '').upper()
                if ('IF' in f_norm and f'D{row}<0' in f_norm and
                        'OVERCAPACITY' in f_norm.replace('"', '').replace("'", '')):
                    ca_checks['F_formula'] += 1
                else:
                    print(f"FAIL: CapacityAnalysis F{row} unexpected formula: {repr(f_val)}")
            else:
                print(f"FAIL: CapacityAnalysis F{row} not a formula: {repr(f_val)}")

        # Score based on how many sub-checks passed
        total_sub = sum(ca_checks.values())
        max_sub = ca_total_per * 5  # 8 rows * 5 column types = 40

        if total_sub == max_sub:
            print(f"PASS: Component 3 — All CapacityAnalysis formulas correct (B-F, rows 2-9) (0.30 pts)")
            total_score += 0.30
        elif total_sub >= max_sub // 2:
            partial = round(0.30 * (total_sub / max_sub), 4)
            print(f"PARTIAL: Component 3 — {total_sub}/{max_sub} CapacityAnalysis formula checks passed (+{partial} pts)")
            print(f"  Sub-checks: {ca_checks}")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {total_sub}/{max_sub} CapacityAnalysis formula checks passed")
            print(f"  Sub-checks: {ca_checks}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: Red conditional formatting on CapacityAnalysis rows (0.20 pts)
    # Task says to "highlight any work center that is over capacity"
    # Golden file has a formula-based conditional format: $D2<0 → red fill (FFFF0000)
    # Initial file has no conditional formatting
    # -------------------------------------------------------------------------
    try:
        cf_found = False
        cf_correct_range = False
        cf_correct_formula = False
        cf_correct_color = False

        for cf_range in ws_ca.conditional_formatting:
            cf_str = str(cf_range)
            rules = ws_ca.conditional_formatting[cf_range]
            for rule in rules:
                if hasattr(rule, 'formula') and rule.formula:
                    for formula in rule.formula:
                        formula_norm = str(formula).replace(' ', '').upper()
                        # Accept formulas like $D2<0 or D2<0
                        if '<0' in formula_norm and 'D' in formula_norm:
                            cf_found = True
                            # Check range covers rows 2-9 across some columns
                            if '2' in cf_str and ('9' in cf_str or 'F' in cf_str.upper()):
                                cf_correct_range = True
                            cf_correct_formula = True
                            # Check fill color
                            try:
                                if rule.dxf and rule.dxf.fill:
                                    color = rule.dxf.fill.fgColor.rgb
                                    # Accept red variants: FFFF0000 or FF0000
                                    if 'FF0000' in color.upper():
                                        cf_correct_color = True
                                        print(f"  CF color found: {color}")
                                    else:
                                        print(f"  CF color wrong: expected red (FFFF0000), got {color}")
                            except Exception as ce:
                                print(f"  CF color check error: {ce}")

        if cf_found and cf_correct_formula and cf_correct_color and cf_correct_range:
            print(f"PASS: Component 4 — Red conditional formatting for OVER CAPACITY rows found (0.20 pts)")
            total_score += 0.20
        elif cf_found and cf_correct_formula and cf_correct_color:
            # Has the right formula and color but range may differ
            print(f"PARTIAL: Component 4 — Conditional formatting found with correct formula and color but range check failed (+0.15 pts)")
            total_score += 0.15
        elif cf_found and cf_correct_formula:
            print(f"PARTIAL: Component 4 — Conditional formatting formula found but color is wrong (+0.10 pts)")
            total_score += 0.10
        elif cf_found:
            print(f"PARTIAL: Component 4 — Conditional formatting found but formula/color not matching task requirements (+0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4 — No conditional formatting found in CapacityAnalysis sheet for OVER CAPACITY highlighting")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
