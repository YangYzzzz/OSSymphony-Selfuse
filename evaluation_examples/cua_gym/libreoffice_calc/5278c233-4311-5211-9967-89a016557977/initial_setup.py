"""
Initial Setup: Credit card transaction data cleanup task
Task ID: calc_gen_data_cleanup_043
Domain: libreoffice_calc

Creates 'CCtransactions' sheet with 150 transactions:
- Column A: Date
- Column B: Description (all uppercase, with SQ *, PAYPAL *, TST*, etc. prefixes)
- Column C: Amount
- Column D: Category (empty — to be filled by agent using IF/ISNUMBER/SEARCH)
- Column E: Empty (agent will add Clean Description formulas here)
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date, timedelta
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_data_cleanup_043'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'CCtransactions'

    # --- Headers ---
    headers = ['Date', 'Description', 'Amount', 'Category']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # --- Transaction data ---
    # Pools of realistic merchant descriptions (uppercase, with messy prefixes)
    dining_raw = [
        'SQ *COFFEESHOP DOWNTOWN', 'TST*PIZZA PALACE', 'SQ *BLUE BOTTLE COFFEE',
        'PAYPAL *FOODPANDA', 'TST*THE SUSHI BAR', 'SQ *BREAKFAST NOOK',
        'AMZN MKTP US*RESTAURANT', 'SQ *MORNING BREW CAFE', 'TST*CHICAGO DEEP DISH',
        'PAYPAL *VENDOR CATERING', 'SQ *TACO STAND CENTRAL', 'TST*BISTRO FRENCH',
        'SQ *BURGER JOINT EXPRESS', 'PAYPAL *DINING VOUCHER', 'TST*THAI KITCHEN',
        'SQ *COASTAL SEAFOOD', 'TST*PIZZA HIVE', 'SQ *NOODLE HOUSE',
        'PAYPAL *BAKERY DELIGHT', 'TST*WINGS AND THINGS',
    ]
    shopping_raw = [
        'AMZN MKTP US*2G4K7', 'AMAZON.COM*AB8C2', 'TARGET 00124 STORE',
        'WALMART SUPERCENTER', 'AMZN MKTP US*9X3M1', 'TARGET CORPORATION',
        'AMAZON PRIME*MONTHLY', 'WALMART.COM ONLINE', 'AMZN MKTP US*7H2P4',
        'TARGET OPTICAL 089', 'AMZN MKTP US*3Y5L9', 'WALMART NEIGHBORHOOD',
        'AMAZON WEB SERVICES', 'AMZN MKTP US*5Q1R8', 'TARGET STORE 0234',
        'WALMART GAS STATION', 'AMZN MKTP US*1D6N3', 'TARGET PICKUP ORDER',
        'AMAZON FRESH ORDERS', 'WALMART VISION CTR',
    ]
    transport_raw = [
        'UBER *TRIP XYZABC', 'LYFT *RIDE 123DEF', 'SHELL OIL 87234',
        'BP GAS STATION 445', 'UBER EATS *ORDER', 'LYFT RIDES INC',
        'SHELL GAS & CONVENIENCE', 'UBER *TRIPFEE 98QRS',
        'BP CONNECT 2234', 'LYFT *RIDE 456GHI', 'UBER *TRIP ABCDEF',
        'GAS STOP SHELL FUELS', 'BP AMOCO 77834', 'UBER CASH PROMO',
        'LYFT *RIDE ABC789', 'SHELL FUEL REWARDS', 'UBER *TRIP 56JKLM',
        'LYFT LINK PROGRAM', 'BP ENERGY 88112', 'UBER RIDES INC',
    ]
    other_raw = [
        'PAYPAL *FREELANCE PMT', 'VENMO PAYMENT RCVD', 'NETFLIX.COM MONTHLY',
        'SPOTIFY USA PREMIUM', 'APPLE.COM/BILL SVCS', 'GOOGLE *GSUITE',
        'PAYPAL *PERSONAL XFER', 'MICROSOFT *OFFICE365', 'HULU SUBSCRIPTION',
        'PAYPAL *CHARITY ORG', 'ZOOM.US MONTHLY PLN', 'DROPBOX INC STORAGE',
        'PAYPAL *EBAY PAYMENT', 'ADOBE CREATIVE CLD', 'GITHUB INC MONTHLY',
        'PAYPAL *SHOPIFY ORG', 'SLACK TECHNOLOGIES', 'CANVA PTY LTD',
        'PAYPAL *FIVERR LLC', 'NOTION LABS INC',
    ]

    all_transactions = []
    start_date = date(2024, 7, 1)
    categories_pool = (
        [(d, 'dining') for d in dining_raw] * 4 +
        [(s, 'shopping') for s in shopping_raw] * 4 +
        [(t, 'transport') for t in transport_raw] * 4 +
        [(o, 'other') for o in other_raw] * 4
    )
    random.shuffle(categories_pool)
    # Use first 150
    for i in range(150):
        desc, _ = categories_pool[i]
        tx_date = start_date + timedelta(days=random.randint(0, 245))
        # Amounts: most positive (purchases), a few negative (refunds)
        if random.random() < 0.08:
            amount = round(-random.uniform(5, 80), 2)
        else:
            amount = round(random.uniform(4.5, 320.0), 2)
        all_transactions.append((tx_date, desc, amount))

    # Sort by date
    all_transactions.sort(key=lambda x: x[0])

    for row_idx, (tx_date, desc, amount) in enumerate(all_transactions, 2):
        ws.cell(row=row_idx, column=1, value=tx_date).number_format = 'yyyy-mm-dd'
        ws.cell(row=row_idx, column=2, value=desc)
        ws.cell(row=row_idx, column=3, value=amount).number_format = '$#,##0.00'
        # Column D (Category) — EMPTY
        # Column E — EMPTY

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 36

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: CCtransactions')
    print(f'  Rows: 151 (1 header + 150 data rows)')
    print(f'  Columns A-D populated, E empty, D column values empty')

create_initial()
