"""
Reward Script: Clean up credit card transaction descriptions and add category column
Task ID: calc_gen_data_cleanup_043
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Column E header is 'Clean Description'              (0.10 pts)
  Component 2: E2:E151 contain PROPER+SUBSTITUTE clean formulas    (0.35 pts)
  Component 3: D2:D151 contain IF/ISNUMBER/SEARCH category formula (0.35 pts)
  Component 4: Conditional formatting on D column (4 category rules)(0.20 pts)
  Total:                                                            (1.00 pts)

Key verification strategy:
- Initial file: D empty, E empty/absent — all 4 components must fail on initial
- Golden file: D has IF/ISNUMBER/SEARCH formulas, E has PROPER/SUBSTITUTE formulas,
  E1='Clean Description', conditional formatting on D
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — reward scripts run on the VM
TASK_ID = 'calc_gen_data_cleanup_043'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: check sheet exists
    if 'CCtransactions' not in wb.sheetnames:
        print("CRITICAL: Sheet 'CCtransactions' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['CCtransactions']

    # Component 1: Column E header is 'Clean Description' (0.10 points)
    # This FAILS on initial (E1 is None) and PASSES on golden (E1='Clean Description')
    try:
        e1_value = ws.cell(row=1, column=5).value
        if e1_value is not None and str(e1_value).strip().lower() == 'clean description':
            print(f"PASS: Component 1 — E1 header is 'Clean Description' (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 — Expected E1='Clean Description', found: {repr(e1_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: E2:E151 contain PROPER+SUBSTITUTE formula for clean descriptions (0.35 points)
    # This FAILS on initial (E2:E151 are all None) and PASSES on golden (all have formulas)
    # Checking for:
    #   - At least 90% of E2:E151 have a formula
    #   - Formulas use PROPER and at least one SUBSTITUTE
    #   - Handle prefix stripping (SQ *, PAYPAL *, TST* or similar)
    try:
        formula_count = 0
        proper_sub_count = 0
        total_rows = 150  # rows 2-151

        for row in range(2, 152):
            val = ws.cell(row=row, column=5).value
            if isinstance(val, str) and val.startswith('='):
                formula_count += 1
                val_upper = val.upper()
                # Must use PROPER for title case conversion
                # Must use SUBSTITUTE to strip prefix patterns
                if 'PROPER' in val_upper and 'SUBSTITUTE' in val_upper:
                    proper_sub_count += 1

        coverage_ratio = formula_count / total_rows
        proper_sub_ratio = proper_sub_count / total_rows

        if coverage_ratio >= 0.90 and proper_sub_ratio >= 0.85:
            print(f"PASS: Component 2 — {formula_count}/{total_rows} E-column cells have PROPER/SUBSTITUTE formula "
                  f"({proper_sub_count} with correct pattern) (0.35 pts)")
            total_score += 0.35
        elif coverage_ratio >= 0.50:
            # Partial credit: formulas exist but may be incomplete
            partial = round(0.35 * (coverage_ratio * 0.5 + proper_sub_ratio * 0.5), 4)
            print(f"PARTIAL: Component 2 — {formula_count}/{total_rows} cells have formulas, "
                  f"{proper_sub_count} use PROPER+SUBSTITUTE. Partial score: {partial}")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {formula_count}/{total_rows} E-column cells have formulas "
                  f"(need >= 90%)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: D2:D151 contain IF/ISNUMBER/SEARCH categorization formula (0.35 points)
    # This FAILS on initial (D2:D151 are all None/empty) and PASSES on golden
    # Checking for:
    #   - At least 90% of D2:D151 have a formula
    #   - Formulas use IF, ISNUMBER, and SEARCH
    #   - Formulas reference at least these categories: Dining, Shopping, Transport, Other
    try:
        d_formula_count = 0
        d_correct_formula_count = 0
        total_rows = 150

        for row in range(2, 152):
            val = ws.cell(row=row, column=4).value
            if isinstance(val, str) and val.startswith('='):
                d_formula_count += 1
                val_upper = val.upper()
                # Formula should use IF, ISNUMBER, SEARCH for auto-categorization
                # and reference the required category outputs
                has_if = 'IF(' in val_upper or 'IF (' in val_upper
                has_isnumber = 'ISNUMBER' in val_upper
                has_search = 'SEARCH' in val_upper
                has_categories = ('"DINING"' in val_upper or '"Dining"' in val or
                                  '"SHOPPING"' in val_upper or '"Shopping"' in val or
                                  '"TRANSPORT"' in val_upper or '"Transport"' in val or
                                  '"OTHER"' in val_upper or '"Other"' in val)
                if has_if and has_isnumber and has_search and has_categories:
                    d_correct_formula_count += 1

        coverage_ratio = d_formula_count / total_rows
        correct_ratio = d_correct_formula_count / total_rows

        if coverage_ratio >= 0.90 and correct_ratio >= 0.85:
            print(f"PASS: Component 3 — {d_formula_count}/{total_rows} D-column cells have IF/ISNUMBER/SEARCH "
                  f"categorization formula ({d_correct_formula_count} with all required elements) (0.35 pts)")
            total_score += 0.35
        elif coverage_ratio >= 0.50:
            partial = round(0.35 * (coverage_ratio * 0.5 + correct_ratio * 0.5), 4)
            print(f"PARTIAL: Component 3 — {d_formula_count}/{total_rows} D cells have formulas, "
                  f"{d_correct_formula_count} use IF+ISNUMBER+SEARCH. Partial: {partial}")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {d_formula_count}/{total_rows} D-column cells have formulas "
                  f"(need >= 90%)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Conditional formatting on D column with category-based rules (0.20 points)
    # This FAILS on initial (no conditional formatting) and PASSES on golden
    # Checking for:
    #   - At least one conditional formatting range covering D column
    #   - At least 3 rules (for Dining, Shopping, Transport, Other or similar)
    try:
        cf_rules = ws.conditional_formatting
        d_cf_ranges = []
        d_cf_rule_count = 0

        for cf_range, rules in cf_rules._cf_rules.items():
            range_str = str(cf_range)
            # Check if the range covers column D (column 4)
            if 'D' in range_str.upper():
                d_cf_ranges.append(range_str)
                d_cf_rule_count += len(rules)

        if d_cf_ranges and d_cf_rule_count >= 3:
            print(f"PASS: Component 4 — Conditional formatting on D column found: "
                  f"{len(d_cf_ranges)} range(s), {d_cf_rule_count} rules (0.20 pts)")
            total_score += 0.20
        elif d_cf_ranges and d_cf_rule_count >= 1:
            # Partial credit: some rules but not enough
            partial = round(0.20 * (d_cf_rule_count / 4), 4)
            print(f"PARTIAL: Component 4 — {d_cf_rule_count} conditional formatting rule(s) on D column "
                  f"(expected >= 3). Partial: {partial}")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — No conditional formatting found on D column. "
                  f"D CF ranges: {d_cf_ranges}, rule count: {d_cf_rule_count}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
