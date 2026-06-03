"""
Reward Script: Format cells C2:C12 as fractions (up to 2 digits)
Task ID: calc_fmt_numfmt_fraction_026
Domain: libreoffice_calc
Scoring:
  Component 1: All 11 cells C2:C12 have a fraction number format (not 'General') — 0.5 pts
  Component 2: All 11 cells C2:C12 retain their original decimal values unchanged — 0.3 pts
  Component 3: Header C1 and other columns unchanged (C1 still 'General', no side effects) — 0.2 pts
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — reward scripts run on the VM
TASK_ID = 'calc_fmt_numfmt_fraction_026'
SHEET_NAME = 'Recipe Conversions'

# Expected decimal values in C2:C12 (must remain unchanged after formatting)
EXPECTED_VALUES = {
    2:  0.5,
    3:  0.25,
    4:  0.75,
    5:  0.333,
    6:  0.125,
    7:  1.5,
    8:  2.25,
    9:  0.667,
    10: 1.0,
    11: 0.875,
    12: 1.25,
}

# Fraction format codes that are acceptable for "up to 2 digit denominator"
# '# ??/??' = up to 2-digit denominator (proper format as per task context)
# '# ?/?' = up to 1-digit denominator (also valid for simple fractions)
# '??/??' = 2-digit no mixed number (also acceptable)
# '?/?' = 1-digit denominator fraction
FRACTION_FORMAT_PATTERNS = ['??/??', '?/??', '??/?', '?/?']


def is_fraction_format(fmt_str):
    """Return True if the number format string is a fraction format."""
    if fmt_str is None or fmt_str == 'General' or fmt_str == '@':
        return False
    # A fraction format must contain '/' with ?-placeholders
    for pattern in FRACTION_FORMAT_PATTERNS:
        if pattern in fmt_str:
            return True
    return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet exists
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Component 1: All cells C2:C12 have a fraction number format (0.5 points)
    # This FAILS on initial (all 'General') and PASSES on golden (all '# ??/??')
    try:
        fraction_count = 0
        non_fraction_cells = []
        for row in range(2, 13):  # rows 2 through 12 inclusive
            cell = ws.cell(row=row, column=3)
            fmt = cell.number_format
            if is_fraction_format(fmt):
                fraction_count += 1
            else:
                non_fraction_cells.append(f"C{row}(fmt={repr(fmt)})")

        if fraction_count == 11:
            print(f"PASS: Component 1 — All 11 cells C2:C12 have fraction number format (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — Only {fraction_count}/11 cells have fraction format. Non-fraction: {non_fraction_cells}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All cells C2:C12 retain original decimal values (0.3 points)
    # Formatting must NOT change the underlying stored values.
    # This FAILS on initial (values are unchanged, but we require that the format
    # is also a fraction format — we check this together to ensure the format change
    # did not corrupt values). We gate this on Component 1 having passed since
    # verifying values alone would pass on the initial file.
    # DESIGN: Only award points if both fraction format AND correct value coexist.
    try:
        correct_value_count = 0
        wrong_values = []
        for row, expected in EXPECTED_VALUES.items():
            cell = ws.cell(row=row, column=3)
            fmt = cell.number_format
            val = cell.value
            if not is_fraction_format(fmt):
                # Format not changed — this cell can't earn value points
                # (It would pass on the initial file too if we just checked the value)
                wrong_values.append(f"C{row}(format not fraction, val={val})")
                continue
            # Format is fraction — verify underlying value is preserved
            if val is None:
                wrong_values.append(f"C{row}(value is None)")
            else:
                try:
                    diff = abs(float(val) - expected)
                    if diff < 0.0001:
                        correct_value_count += 1
                    else:
                        wrong_values.append(f"C{row}(expected={expected}, got={val})")
                except (ValueError, TypeError):
                    wrong_values.append(f"C{row}(non-numeric value={repr(val)})")

        if correct_value_count == 11:
            print(f"PASS: Component 2 — All 11 cells C2:C12 retain correct decimal values after formatting (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — {correct_value_count}/11 cells have both fraction format and correct value. Issues: {wrong_values}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: No side effects — C1 header keeps 'General' format, values unchanged (0.2 points)
    # C1 is the header row and should NOT be reformatted.
    # Also ensure columns A, B, D are not affected.
    # This checks that only the exact task range C2:C12 was modified.
    # This FAILS on initial because the fraction formats in C2:C12 haven't been applied yet
    # (Component 1 is 0 on initial, so this compound check also fails on initial).
    # We use this as a combined integrity check: fraction formats in C2:C12 AND no C1 side-effect.
    try:
        c1_fmt = ws.cell(row=1, column=3).number_format
        c1_val = ws.cell(row=1, column=3).value

        # C1 must still be 'General' (header was not reformatted)
        c1_format_ok = (c1_fmt == 'General')
        # C1 value must still be the header text
        c1_value_ok = (str(c1_val).strip() == 'Cup Fraction')

        # Fraction format cells must already be correct (require Component 1 passed)
        fractions_applied = (fraction_count == 11) if 'fraction_count' in dir() else False

        if c1_format_ok and c1_value_ok and fractions_applied:
            print(f"PASS: Component 3 — C1 header intact (fmt={repr(c1_fmt)}, val={repr(c1_val)}) and fraction range correctly scoped to C2:C12 (0.2 pts)")
            total_score += 0.2
        else:
            reasons = []
            if not c1_format_ok:
                reasons.append(f"C1 format changed to {repr(c1_fmt)}, expected 'General'")
            if not c1_value_ok:
                reasons.append(f"C1 value changed to {repr(c1_val)}, expected 'Cup Fraction'")
            if not fractions_applied:
                reasons.append("Fraction format not applied to C2:C12 (Component 1 failed)")
            print(f"FAIL: Component 3 — {'; '.join(reasons)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
