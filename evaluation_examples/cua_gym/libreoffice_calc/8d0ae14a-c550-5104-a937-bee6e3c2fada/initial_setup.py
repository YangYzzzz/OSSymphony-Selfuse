"""
Initial Setup: Recipe Conversions spreadsheet with decimal values in column C
Task ID: calc_fmt_numfmt_fraction_026
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_numfmt_fraction_026'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Recipe Conversions ---
    ws = wb.active
    ws.title = 'Recipe Conversions'

    # Headers
    headers = ['Ingredient', 'Metric Amount', 'Cup Fraction', 'Notes']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Recipe data rows 2-12
    # Column A: Ingredient, B: Metric Amount (grams), C: decimal fraction value, D: Notes
    data = [
        ['All-purpose flour',    120,  0.5,   'Standard sifted measure'],
        ['Granulated sugar',      50,  0.25,  'Fine white sugar'],
        ['Unsalted butter',      170,  0.75,  'Room temperature'],
        ['Milk (whole)',          80,  0.333, 'Full-fat recommended'],
        ['Baking powder',         15,  0.125, 'Leveling agent'],
        ['Vegetable oil',        355,  1.5,   'Neutral flavor'],
        ['Brown sugar (packed)', 450,  2.25,  'Firmly packed'],
        ['Heavy cream',          160,  0.667, 'Whipping cream'],
        ['Honey',                340,  1.0,   'Raw or filtered'],
        ['Almond flour',         200,  0.875, 'Blanched preferred'],
        ['Rolled oats',          240,  1.25,  'Old-fashioned style'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Column C uses General format (no fraction format) - this is the initial state
    # The task is to apply fraction format to C2:C12
    for row in range(2, 13):
        ws.cell(row=row, column=3).number_format = 'General'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 26

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
