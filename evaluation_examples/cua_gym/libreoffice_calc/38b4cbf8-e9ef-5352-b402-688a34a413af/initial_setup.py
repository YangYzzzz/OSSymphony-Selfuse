"""
Initial Setup: Logistics shipment dataset with blank Project Codes in column A.
Task ID: osworld_calc_fill_blanks_above_006
Domain: libreoffice_calc

Creates a spreadsheet with:
  - Column A: Project Code (some cells intentionally blank — agent must fill from above)
  - Column B: Shipment ID
  - Column C: Weight kg
  - Column D: Date
  - Column E: Empty (agent must add running-total formulas)
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_fill_blanks_above_006'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shipments"

    # Headers
    headers = ['Project Code', 'Shipment ID', 'Weight kg', 'Date']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Data rows — Project Code (col A) has blanks where the same project continues.
    # Three projects: PROJ-ALPHA, PROJ-BETA, PROJ-GAMMA
    # Blanks are represented as None — the agent must fill them from the row above.
    data = [
        # Project PROJ-ALPHA (rows 2-6)
        ['PROJ-ALPHA', 'SHP-1001', 124.5,  '2025-01-05'],
        [None,         'SHP-1002',  88.0,  '2025-01-06'],
        [None,         'SHP-1003', 213.0,  '2025-01-08'],
        [None,         'SHP-1004',  55.5,  '2025-01-10'],
        [None,         'SHP-1005', 177.0,  '2025-01-11'],
        # Project PROJ-BETA (rows 7-11)
        ['PROJ-BETA',  'SHP-1006', 340.0,  '2025-01-14'],
        [None,         'SHP-1007', 102.5,  '2025-01-15'],
        [None,         'SHP-1008',  67.0,  '2025-01-16'],
        [None,         'SHP-1009', 289.5,  '2025-01-17'],
        [None,         'SHP-1010', 155.0,  '2025-01-20'],
        # Project PROJ-GAMMA (rows 12-16)
        ['PROJ-GAMMA', 'SHP-1011', 198.0,  '2025-01-21'],
        [None,         'SHP-1012',  73.5,  '2025-01-22'],
        [None,         'SHP-1013', 310.0,  '2025-01-23'],
        [None,         'SHP-1014', 445.0,  '2025-01-24'],
        [None,         'SHP-1015',  92.0,  '2025-01-27'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Column E is intentionally empty — no header, no formulas

    # Set column widths for readability
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
