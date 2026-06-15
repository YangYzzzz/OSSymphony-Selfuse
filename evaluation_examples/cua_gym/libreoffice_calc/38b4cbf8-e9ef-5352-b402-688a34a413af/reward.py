"""
Reward Script: Fill blanks in column A and add running total column
Task ID: osworld_calc_fill_blanks_above_006
Domain: libreoffice_calc
Scoring:
  Component 1: All blank cells in column A are filled with the project code from the row above (0.4 pts)
  Component 2: Column E has a header for the running total (0.1 pts)
  Component 3: Column E rows 2-16 contain running-total formulas (cumulative weight per project) (0.5 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_fill_blanks_above_006'

# Expected blank rows in column A (1-indexed) in the INITIAL file
# These rows had None and should now be filled with the project code from above
EXPECTED_BLANK_ROWS = [3, 4, 5, 6, 8, 9, 10, 11, 13, 14, 15, 16]

# Expected project code assignments after fill-down
EXPECTED_COL_A = {
    2: 'PROJ-ALPHA',
    3: 'PROJ-ALPHA',
    4: 'PROJ-ALPHA',
    5: 'PROJ-ALPHA',
    6: 'PROJ-ALPHA',
    7: 'PROJ-BETA',
    8: 'PROJ-BETA',
    9: 'PROJ-BETA',
    10: 'PROJ-BETA',
    11: 'PROJ-BETA',
    12: 'PROJ-GAMMA',
    13: 'PROJ-GAMMA',
    14: 'PROJ-GAMMA',
    15: 'PROJ-GAMMA',
    16: 'PROJ-GAMMA',
}

DATA_ROWS = list(range(2, 17))  # rows 2 to 16 (15 data rows)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists
    if 'Shipments' not in wb.sheetnames:
        print("CRITICAL: 'Shipments' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Shipments']

    # -------------------------------------------------------------------
    # Component 1: All blank cells in column A are now filled (0.4 points)
    # The initial file had 12 blank cells in col A (rows 3-6, 8-11, 13-16).
    # Each must now be filled with the correct project code from above.
    # -------------------------------------------------------------------
    try:
        filled_correct = 0
        filled_total = 0
        errors_comp1 = []

        for row, expected_code in EXPECTED_COL_A.items():
            val = ws.cell(row=row, column=1).value
            filled_total += 1
            if val is not None and str(val).strip() == expected_code:
                filled_correct += 1
            else:
                errors_comp1.append(f"Row {row}: expected '{expected_code}', found {repr(val)}")

        # Give points proportionally but only award full 0.4 if all blanks correctly filled
        if filled_correct == filled_total:
            total_score += 0.4
            print(f"PASS: Component 1 — All {filled_correct}/{filled_total} col A cells correctly filled (0.4 pts)")
        elif filled_correct >= filled_total * 0.5:
            partial = round(0.4 * (filled_correct / filled_total), 2)
            total_score += partial
            print(f"PARTIAL: Component 1 — {filled_correct}/{filled_total} col A cells correctly filled ({partial} pts)")
            print(f"  Issues: {errors_comp1[:3]}")
        else:
            print(f"FAIL: Component 1 — Only {filled_correct}/{filled_total} col A cells correctly filled (0.0 pts)")
            print(f"  Issues: {errors_comp1[:3]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------
    # Component 2: Column E header exists for the running total (0.1 points)
    # The initial file had only 4 columns; the golden adds column E.
    # -------------------------------------------------------------------
    try:
        e1_val = ws.cell(row=1, column=5).value
        if e1_val is not None and str(e1_val).strip() != '':
            print(f"PASS: Component 2 — Column E header present: '{e1_val}' (0.1 pts)")
            total_score += 0.1
        else:
            print(f"FAIL: Component 2 — Column E header missing or empty (found: {repr(e1_val)})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------
    # Component 3: Column E rows 2-16 have running-total formulas (0.5 points)
    # Formulas must implement cumulative weight per project, resetting at
    # each new project. Valid approaches:
    #   - IF-based: =IF(Ax=A(x-1), E(x-1)+Cx, Cx)  or SUMIF-based
    #   - First row: =C2 (just the weight, no previous row)
    # We check:
    #   (a) All 15 cells in E2:E16 are non-empty
    #   (b) The formulas reference column C (weight) and column A (project)
    #       OR the cell contains numeric values consistent with running totals
    # -------------------------------------------------------------------
    try:
        formula_present = 0
        formula_correct_logic = 0
        formula_errors = []

        for row in DATA_ROWS:
            val = ws.cell(row=row, column=5).value
            if val is None:
                formula_errors.append(f"Row {row}: E{row} is empty")
                continue

            formula_present += 1
            val_str = str(val).strip().upper().replace(' ', '')

            if row == 2:
                # First data row: should just be =C2 or the numeric value of C2
                # Accept: formula referencing C2, or the numeric weight itself
                if isinstance(val, (int, float)):
                    # Numeric value stored (no formula) — check it equals C2's weight
                    c2_weight = ws.cell(row=2, column=3).value
                    if c2_weight is not None and abs(float(val) - float(c2_weight)) < 0.01:
                        formula_correct_logic += 1
                    else:
                        formula_errors.append(f"Row 2: E2={val} does not match C2 weight ({c2_weight})")
                elif 'C2' in val_str:
                    # Formula references C2
                    formula_correct_logic += 1
                else:
                    formula_errors.append(f"Row 2: E2 formula '{val}' does not reference C2")
            else:
                # Rows 3-16: should have running total logic
                # Accept IF-based or SUMIF-based formulas referencing A and C columns
                # Or numeric values consistent with cumulative weight
                if isinstance(val, str) and val.startswith('='):
                    # Check formula references both column A (project check) and column C (weight)
                    has_col_a = 'A' in val_str
                    has_col_c = 'C' in val_str
                    has_prev_e = 'E' in val_str or 'SUMIF' in val_str
                    if has_col_a and has_col_c:
                        formula_correct_logic += 1
                    elif has_col_c and has_prev_e:
                        # SUMIF-based approach also valid
                        formula_correct_logic += 1
                    else:
                        formula_errors.append(f"Row {row}: E{row}='{val}' missing expected columns A/C reference")
                elif isinstance(val, (int, float)):
                    # Numeric value — treat as possibly valid (cached value from formula)
                    formula_correct_logic += 1
                else:
                    formula_errors.append(f"Row {row}: E{row}='{val}' unexpected type")

        total_cells = len(DATA_ROWS)  # 15

        if formula_present == total_cells and formula_correct_logic == total_cells:
            total_score += 0.5
            print(f"PASS: Component 3 — All {total_cells} running total formulas present and correct (0.5 pts)")
        elif formula_present == total_cells and formula_correct_logic >= total_cells * 0.8:
            partial = round(0.5 * (formula_correct_logic / total_cells), 2)
            total_score += partial
            print(f"PARTIAL: Component 3 — {formula_correct_logic}/{total_cells} formulas correct ({partial} pts)")
            if formula_errors:
                print(f"  Issues: {formula_errors[:3]}")
        elif formula_present > 0:
            partial = round(0.5 * (formula_present / total_cells) * 0.5, 2)
            total_score += partial
            print(f"PARTIAL: Component 3 — Only {formula_present}/{total_cells} cells populated ({partial} pts)")
            if formula_errors:
                print(f"  Issues: {formula_errors[:3]}")
        else:
            print(f"FAIL: Component 3 — Column E has no running total data (0.0 pts)")
            if formula_errors:
                print(f"  Issues: {formula_errors[:3]}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {round(final_score, 4)}")
    return final_score


# Default: test against canonical artifact path on the VM
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
