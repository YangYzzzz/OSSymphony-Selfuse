"""
Initial Setup: Pivot table layout with Region as rows, Quarter as columns
Task ID: calc_adv_pivot_layout_009
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_adv_pivot_layout_009'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    # ---- Sheet 1: SalesLog (source data) ----
    ws_log = wb.active
    ws_log.title = 'SalesLog'

    headers = ['Region', 'Quarter', 'Sales', 'Rep']
    for col, h in enumerate(headers, 1):
        cell = ws_log.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFFFF', name='Calibri', size=11)
        cell.alignment = Alignment(horizontal='center')

    # Realistic sales data: 4 regions x 4 quarters = 16 rows
    raw_data = [
        ('North', 'Q1', 128450, 'Daniel Reeves'),
        ('North', 'Q2', 143200, 'Daniel Reeves'),
        ('North', 'Q3', 159800, 'Sandra Kowalski'),
        ('North', 'Q4', 175300, 'Sandra Kowalski'),
        ('South', 'Q1', 98700,  'Marcus Johnson'),
        ('South', 'Q2', 112600, 'Marcus Johnson'),
        ('South', 'Q3', 105400, 'Priya Sharma'),
        ('South', 'Q4', 134200, 'Priya Sharma'),
        ('East',  'Q1', 164300, 'Yuki Tanaka'),
        ('East',  'Q2', 178900, 'Yuki Tanaka'),
        ('East',  'Q3', 192500, 'Carlos Mendez'),
        ('East',  'Q4', 210400, 'Carlos Mendez'),
        ('West',  'Q1', 145600, 'Sarah Chen'),
        ('West',  'Q2', 162300, 'Sarah Chen'),
        ('West',  'Q3', 171800, 'Rachel Torres'),
        ('West',  'Q4', 188700, 'Rachel Torres'),
    ]

    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r, row_data in enumerate(raw_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_log.cell(row=r, column=c, value=val)
            cell.border = border
            if c == 3:  # Sales column
                cell.number_format = '#,##0'

    # Set column widths
    ws_log.column_dimensions['A'].width = 12
    ws_log.column_dimensions['B'].width = 10
    ws_log.column_dimensions['C'].width = 14
    ws_log.column_dimensions['D'].width = 20
    ws_log.freeze_panes = 'A2'

    # ---- Sheet 2: PivotView (pivot table layout) ----
    # Current layout: Region as rows, Quarter as columns
    ws_pv = wb.create_sheet('PivotView')

    # Header cell (top-left corner)
    corner = ws_pv.cell(row=1, column=1, value='Region \\ Quarter')
    corner.font = Font(bold=True, name='Calibri', size=11, color='FFFFFFFF')
    corner.fill = PatternFill(start_color='FF2E4057', end_color='FF2E4057', fill_type='solid')
    corner.alignment = Alignment(horizontal='center', vertical='center')
    corner.border = border

    # Column headers: Q1, Q2, Q3, Q4, Total
    quarters = ['Q1', 'Q2', 'Q3', 'Q4', 'Total']
    for c, q in enumerate(quarters, 2):
        cell = ws_pv.cell(row=1, column=c, value=q)
        cell.font = Font(bold=True, name='Calibri', size=11, color='FFFFFFFF')
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Row labels: North, South, East, West, Total
    regions = ['North', 'South', 'East', 'West', 'Total']
    region_sales = {
        'North': {'Q1': 128450, 'Q2': 143200, 'Q3': 159800, 'Q4': 175300},
        'South': {'Q1': 98700,  'Q2': 112600, 'Q3': 105400, 'Q4': 134200},
        'East':  {'Q1': 164300, 'Q2': 178900, 'Q3': 192500, 'Q4': 210400},
        'West':  {'Q1': 145600, 'Q2': 162300, 'Q3': 171800, 'Q4': 188700},
    }

    for r, region in enumerate(regions, 2):
        # Row label cell
        label_cell = ws_pv.cell(row=r, column=1, value=region)
        if region == 'Total':
            label_cell.font = Font(bold=True, name='Calibri', size=11, color='FFFFFFFF')
            label_cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        else:
            label_cell.font = Font(bold=True, name='Calibri', size=11)
            label_cell.fill = PatternFill(start_color='FFDCE6F1', end_color='FFDCE6F1', fill_type='solid')
        label_cell.alignment = Alignment(horizontal='left', vertical='center')
        label_cell.border = border

        # Data cells
        for c, q in enumerate(quarters, 2):
            cell = ws_pv.cell(row=r, column=c)
            if region == 'Total' and q == 'Total':
                total_val = sum(
                    region_sales[reg][qt]
                    for reg in ['North', 'South', 'East', 'West']
                    for qt in ['Q1', 'Q2', 'Q3', 'Q4']
                )
                cell.value = total_val
                cell.font = Font(bold=True, name='Calibri', size=11, color='FFFFFFFF')
                cell.fill = PatternFill(start_color='FF2E4057', end_color='FF2E4057', fill_type='solid')
            elif region == 'Total':
                # Quarter total (sum of all regions for that quarter)
                total_q = sum(region_sales[reg][q] for reg in ['North', 'South', 'East', 'West'])
                cell.value = total_q
                cell.font = Font(bold=True, name='Calibri', size=11, color='FFFFFFFF')
                cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
            elif q == 'Total':
                # Region total (sum of all quarters for that region)
                total_r = sum(region_sales[region].values())
                cell.value = total_r
                cell.font = Font(bold=True, name='Calibri', size=11)
                cell.fill = PatternFill(start_color='FFDCE6F1', end_color='FFDCE6F1', fill_type='solid')
            else:
                cell.value = region_sales[region][q]
                cell.font = Font(name='Calibri', size=11)
                # Alternate row coloring
                if (r % 2) == 0:
                    cell.fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='FFF2F7FB', end_color='FFF2F7FB', fill_type='solid')
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='right')
            cell.border = border

    # Column widths for PivotView
    ws_pv.column_dimensions['A'].width = 18
    ws_pv.column_dimensions['B'].width = 12
    ws_pv.column_dimensions['C'].width = 12
    ws_pv.column_dimensions['D'].width = 12
    ws_pv.column_dimensions['E'].width = 12
    ws_pv.column_dimensions['F'].width = 12

    # Title above the pivot
    ws_pv.insert_rows(1)
    title_cell = ws_pv.cell(row=1, column=1, value='Sum of Sales by Region and Quarter')
    title_cell.font = Font(bold=True, name='Calibri', size=14)
    title_cell.alignment = Alignment(horizontal='left')
    ws_pv.row_dimensions[1].height = 24
    ws_pv.row_dimensions[2].height = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: SalesLog, PivotView')
    print('PivotView layout: Region (rows) x Quarter (columns)')

create_initial()
