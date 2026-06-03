"""
Reward Script: Rearrange pivot table — swap Region (rows) and Quarter (columns) fields
Task ID: calc_adv_pivot_layout_009
Domain: libreoffice_calc

Scoring Rubric:
  Component 1 (0.30): Row labels are Quarters (Q1,Q2,Q3,Q4) and column headers are Regions (North,South,East,West)
  Component 2 (0.40): Data values are correctly transposed (specific cell spot-checks + grand total)
  Component 3 (0.30): Title and intersection label updated to reflect Quarter/Region swap

Total: 1.00
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_adv_pivot_layout_009'

# Expected values in the golden (transposed) state
EXPECTED_COLUMN_HEADERS = ['North', 'South', 'East', 'West']  # B2:E2
EXPECTED_ROW_LABELS     = ['Q1', 'Q2', 'Q3', 'Q4']            # A3:A6
EXPECTED_INTERSECTION_LABEL = 'Quarter \\ Region'             # A2

# Expected data grid (rows = Q1..Q4, cols = North,South,East,West)
# Row3=Q1: 128450, 98700, 164300, 145600  -> Total 537050
# Row4=Q2: 143200, 112600, 178900, 162300 -> Total 597000
# Row5=Q3: 159800, 105400, 192500, 171800 -> Total 629500
# Row6=Q4: 175300, 134200, 210400, 188700 -> Total 708600
# Row7=Total: 606750, 450900, 746100, 668400 -> Grand Total 2472150
EXPECTED_DATA = {
    (3, 2): 128450, (3, 3): 98700,  (3, 4): 164300, (3, 5): 145600, (3, 6): 537050,
    (4, 2): 143200, (4, 3): 112600, (4, 4): 178900, (4, 5): 162300, (4, 6): 597000,
    (5, 2): 159800, (5, 3): 105400, (5, 4): 192500, (5, 5): 171800, (5, 6): 629500,
    (6, 2): 175300, (6, 3): 134200, (6, 4): 210400, (6, 5): 188700, (6, 6): 708600,
    (7, 2): 606750, (7, 3): 450900, (7, 4): 746100, (7, 5): 668400, (7, 6): 2472150,
}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: PivotView sheet must exist
    if 'PivotView' not in wb.sheetnames:
        print("FAIL: PivotView sheet not found in workbook")
        print("\nScore: 0.0/1.0")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['PivotView']

    # -----------------------------------------------------------------------
    # Component 1: Row labels are Quarters (Q1-Q4), column headers are Regions
    # (0.30 points) — verifies the axis swap took place
    # -----------------------------------------------------------------------
    try:
        # Check column headers (B2:E2 must be North, South, East, West)
        actual_col_headers = [ws.cell(row=2, column=c).value for c in range(2, 6)]
        col_headers_correct = actual_col_headers == EXPECTED_COLUMN_HEADERS

        # Check row labels (A3:A6 must be Q1, Q2, Q3, Q4)
        actual_row_labels = [ws.cell(row=r, column=1).value for r in range(3, 7)]
        row_labels_correct = actual_row_labels == EXPECTED_ROW_LABELS

        if col_headers_correct and row_labels_correct:
            print(f"PASS: Component 1 — Column headers={actual_col_headers}, "
                  f"Row labels={actual_row_labels} (0.30 pts)")
            total_score += 0.30
        else:
            if not col_headers_correct:
                print(f"FAIL: Component 1 — Column headers wrong. "
                      f"Expected {EXPECTED_COLUMN_HEADERS}, found {actual_col_headers}")
            if not row_labels_correct:
                print(f"FAIL: Component 1 — Row labels wrong. "
                      f"Expected {EXPECTED_ROW_LABELS}, found {actual_row_labels}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Data values correctly transposed
    # (0.40 points) — spot-checks all data cells plus grand total row/col
    # -----------------------------------------------------------------------
    try:
        mismatches = []
        for (row, col), expected_val in EXPECTED_DATA.items():
            actual_val = ws.cell(row=row, column=col).value
            if actual_val != expected_val:
                from openpyxl.utils import get_column_letter
                coord = f"{get_column_letter(col)}{row}"
                mismatches.append(f"{coord}: expected {expected_val}, got {actual_val}")

        if not mismatches:
            print(f"PASS: Component 2 — All {len(EXPECTED_DATA)} data cells correctly transposed (0.40 pts)")
            total_score += 0.40
        else:
            print(f"FAIL: Component 2 — {len(mismatches)} data cell(s) incorrect:")
            for m in mismatches[:5]:  # show first 5
                print(f"  {m}")
            if len(mismatches) > 5:
                print(f"  ... and {len(mismatches) - 5} more")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Title and intersection label updated to reflect swap
    # (0.30 points) — A1 title and A2 intersection label
    # -----------------------------------------------------------------------
    try:
        actual_title = ws.cell(row=1, column=1).value
        actual_intersection = ws.cell(row=2, column=1).value

        # Title should mention Quarter before Region (indicating the swap)
        title_correct = (
            actual_title is not None and
            'Quarter' in str(actual_title) and
            'Region' in str(actual_title) and
            str(actual_title).index('Quarter') < str(actual_title).index('Region')
        )

        # Intersection label should be 'Quarter \ Region' (Quarter listed first)
        intersection_correct = (
            actual_intersection is not None and
            'Quarter' in str(actual_intersection) and
            'Region' in str(actual_intersection) and
            str(actual_intersection).index('Quarter') < str(actual_intersection).index('Region')
        )

        if title_correct and intersection_correct:
            print(f"PASS: Component 3 — Title='{actual_title}', "
                  f"Intersection='{actual_intersection}' (0.30 pts)")
            total_score += 0.30
        else:
            if not title_correct:
                print(f"FAIL: Component 3 — Title '{actual_title}' should put Quarter before Region")
            if not intersection_correct:
                print(f"FAIL: Component 3 — Intersection label '{actual_intersection}' "
                      f"should put Quarter before Region (e.g., 'Quarter \\ Region')")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
