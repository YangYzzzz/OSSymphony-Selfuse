"""
Initial Setup: Date formatting with TEXT formulas breaking calculations
Task ID: calc_tbl_086
Domain: libreoffice_calc

Column A: Start dates (actual date values)
Column B: TEXT formulas converting dates to text (breaking downstream calcs)
Column C: Formulas doing B-A that return errors because B is text
"""

import os
import shlex
import subprocess
import time
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_086'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Project Timeline ---
    ws = wb.active
    ws.title = 'Project Timeline'

    # Header style
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    # Headers
    headers = ['Start Date', 'End Date', 'Duration (Days)', 'Project', 'Phase', 'Lead', 'Budget ($)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 14

    # Project data - realistic entries
    # Column A gets actual date values
    # Column B gets TEXT formulas (converting to text string - the PROBLEM)
    # Column C gets =B{row}-A{row} (will error because B is text)
    projects = [
        (datetime(2025, 1, 6), datetime(2025, 3, 14), 'Atlas CRM Migration', 'Discovery', 'Priya Sharma', 42000),
        (datetime(2025, 2, 3), datetime(2025, 5, 23), 'Atlas CRM Migration', 'Development', 'Marcus Chen', 128000),
        (datetime(2025, 3, 10), datetime(2025, 4, 18), 'Beacon Analytics Dashboard', 'Design', 'Elena Vasquez', 35000),
        (datetime(2025, 4, 1), datetime(2025, 7, 31), 'Beacon Analytics Dashboard', 'Development', 'Tomasz Kowalski', 95000),
        (datetime(2025, 1, 15), datetime(2025, 2, 28), 'Catalyst Mobile App', 'Prototype', 'Aisha Patel', 28000),
        (datetime(2025, 3, 1), datetime(2025, 6, 15), 'Catalyst Mobile App', 'Alpha Build', 'Jordan Williams', 110000),
        (datetime(2025, 5, 1), datetime(2025, 8, 22), 'Catalyst Mobile App', 'Beta Testing', 'Mei Lin', 45000),
        (datetime(2025, 2, 17), datetime(2025, 3, 28), 'Delta Cloud Infrastructure', 'Assessment', 'Raj Gupta', 22000),
        (datetime(2025, 4, 7), datetime(2025, 9, 12), 'Delta Cloud Infrastructure', 'Migration', 'Sofia Rodriguez', 175000),
        (datetime(2025, 6, 1), datetime(2025, 7, 15), 'Echo Security Audit', 'Penetration Test', 'Nathan Brooks', 38000),
        (datetime(2025, 7, 1), datetime(2025, 8, 30), 'Echo Security Audit', 'Remediation', 'Keiko Tanaka', 52000),
        (datetime(2025, 3, 15), datetime(2025, 5, 10), 'Foxtrot Data Pipeline', 'ETL Design', 'David Okafor', 31000),
    ]

    date_format = 'MM/DD/YYYY'
    data_align = Alignment(horizontal='center', vertical='center')

    for r, (start_dt, end_dt, project, phase, lead, budget) in enumerate(projects, 2):
        # Column A: actual date values with date format
        cell_a = ws.cell(row=r, column=1, value=start_dt)
        cell_a.number_format = 'MM/DD/YYYY'
        cell_a.alignment = data_align
        cell_a.border = thin_border

        # Column B: TEXT formula that converts end_date to text string
        # This is the PROBLEM - TEXT() returns a text string, not a date
        # We put the actual date in a helper column (hidden) and reference it
        # Actually, simpler: put the date value and wrap with TEXT formula
        # We'll store end dates in column H (hidden helper) and TEXT formula in B
        ws.cell(row=r, column=8, value=end_dt)  # Helper column H
        ws.cell(row=r, column=8).number_format = 'YYYY-MM-DD'

        cell_b = ws.cell(row=r, column=2, value=f'=TEXT(H{r},"MM/DD/YYYY")')
        cell_b.alignment = data_align
        cell_b.border = thin_border

        # Column C: =B{r}-A{r} - will error because B is text from TEXT()
        cell_c = ws.cell(row=r, column=3, value=f'=B{r}-A{r}')
        cell_c.alignment = data_align
        cell_c.border = thin_border

        # Column D: Project
        cell_d = ws.cell(row=r, column=4, value=project)
        cell_d.border = thin_border

        # Column E: Phase
        cell_e = ws.cell(row=r, column=5, value=phase)
        cell_e.border = thin_border

        # Column F: Lead
        cell_f = ws.cell(row=r, column=6, value=lead)
        cell_f.border = thin_border

        # Column G: Budget
        cell_g = ws.cell(row=r, column=7, value=budget)
        cell_g.number_format = '$#,##0'
        cell_g.border = thin_border

    # Hide helper column H
    ws.column_dimensions['H'].hidden = True

    # Freeze header row
    ws.freeze_panes = 'A2'

    # --- Sheet 2: Summary ---
    ws2 = wb.create_sheet('Summary')
    ws2['A1'] = 'Project Summary'
    ws2['A1'].font = Font(size=14, bold=True)
    ws2.merge_cells('A1:C1')

    ws2['A3'] = 'Total Projects'
    ws2['B3'] = 5
    ws2['A4'] = 'Total Phases'
    ws2['B4'] = 12
    ws2['A5'] = 'Total Budget'
    ws2['B5'] = '=SUM(\'Project Timeline\'!G2:G13)'
    ws2['B5'].number_format = '$#,##0'
    ws2['A7'] = 'Note: End Date column uses TEXT() formula.'
    ws2['A8'] = 'Duration calculations in Column C are broken.'
    ws2['A7'].font = Font(italic=True, color='FF0000')
    ws2['A8'].font = Font(italic=True, color='FF0000')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
