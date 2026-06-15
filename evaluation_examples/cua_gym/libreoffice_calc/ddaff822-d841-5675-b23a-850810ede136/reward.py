"""
Reward Script: Fix TEXT formula dates in Column B — replace with real dates + MM/DD/YYYY format
Task ID: calc_tbl_086
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Column B cells contain datetime values, not TEXT formulas
  Component 2 (0.3): Column B cells have MM/DD/YYYY number format
  Component 3 (0.3): Column B date values match the expected dates from column H
"""

import os
import datetime

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_086'

# Expected dates for column B (derived from the H column source dates)
EXPECTED_DATES = {
    2: datetime.datetime(2025, 3, 14),
    3: datetime.datetime(2025, 5, 23),
    4: datetime.datetime(2025, 4, 18),
    5: datetime.datetime(2025, 7, 31),
    6: datetime.datetime(2025, 2, 28),
    7: datetime.datetime(2025, 6, 15),
    8: datetime.datetime(2025, 8, 22),
    9: datetime.datetime(2025, 3, 28),
    10: datetime.datetime(2025, 9, 12),
    11: datetime.datetime(2025, 7, 15),
    12: datetime.datetime(2025, 8, 30),
    13: datetime.datetime(2025, 5, 10),
}

DATA_ROWS = list(range(2, 14))  # rows 2-13


def persist_app_state(domain):
    """Try to save any unsaved LibreOffice changes."""
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print(f"PERSIST: ctrl+s sent for {domain}")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Project Timeline']

    # Component 1: Column B cells contain actual datetime values, NOT TEXT formulas (0.4 points)
    # In the initial file, B2:B13 have =TEXT(H2,"MM/DD/YYYY") formulas.
    # In the golden file, they should be real datetime objects.
    try:
        datetime_count = 0
        formula_count = 0
        for row in DATA_ROWS:
            val = ws.cell(row=row, column=2).value
            if isinstance(val, datetime.datetime):
                datetime_count += 1
            elif isinstance(val, str) and val.startswith('='):
                formula_count += 1

        if datetime_count == len(DATA_ROWS):
            print(f"PASS: Component 1 — All {datetime_count} B cells are datetime values (0.4 pts)")
            total_score += 0.4
        elif datetime_count > 0 and formula_count == 0:
            # Partial: some are datetime but not all, and no TEXT formulas remain
            partial = 0.4 * (datetime_count / len(DATA_ROWS))
            print(f"PARTIAL: Component 1 — {datetime_count}/{len(DATA_ROWS)} B cells are datetime (no formulas remain) ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — {datetime_count} datetime, {formula_count} formulas in B column")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Column B cells have MM/DD/YYYY number format (0.3 points)
    # In the initial file, B column number_format is 'General'.
    # In the golden file, it should be 'MM/DD/YYYY'.
    try:
        format_count = 0
        for row in DATA_ROWS:
            nf = ws.cell(row=row, column=2).number_format
            # Accept common MM/DD/YYYY variants
            if nf and 'MM' in nf.upper() and 'DD' in nf.upper() and 'YYYY' in nf.upper():
                format_count += 1

        if format_count == len(DATA_ROWS):
            print(f"PASS: Component 2 — All {format_count} B cells have MM/DD/YYYY format (0.3 pts)")
            total_score += 0.3
        elif format_count > 0:
            partial = 0.3 * (format_count / len(DATA_ROWS))
            print(f"PARTIAL: Component 2 — {format_count}/{len(DATA_ROWS)} B cells have MM/DD/YYYY format ({partial:.2f} pts)")
            total_score += partial
        else:
            # Check what format they actually have
            sample_nf = ws.cell(row=2, column=2).number_format
            print(f"FAIL: Component 2 — B cells do not have MM/DD/YYYY format (sample: {sample_nf!r})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Column B date values match expected dates from H column (0.3 points)
    # Only scores if B cells are actually datetime (otherwise Component 1 already failed).
    # This ensures the correct date values were preserved during the fix.
    try:
        match_count = 0
        checked = 0
        for row in DATA_ROWS:
            b_val = ws.cell(row=row, column=2).value
            expected = EXPECTED_DATES[row]
            if isinstance(b_val, datetime.datetime):
                checked += 1
                # Compare date portion only (ignore time component)
                if b_val.date() == expected.date():
                    match_count += 1
                else:
                    print(f"  Row {row}: B={b_val.date()}, expected={expected.date()}")

        if checked == 0:
            print(f"FAIL: Component 3 — No datetime values in B to compare")
        elif match_count == len(DATA_ROWS):
            print(f"PASS: Component 3 — All {match_count} B dates match expected values (0.3 pts)")
            total_score += 0.3
        elif match_count > 0:
            partial = 0.3 * (match_count / len(DATA_ROWS))
            print(f"PARTIAL: Component 3 — {match_count}/{len(DATA_ROWS)} dates match ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No dates match (checked {checked})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
