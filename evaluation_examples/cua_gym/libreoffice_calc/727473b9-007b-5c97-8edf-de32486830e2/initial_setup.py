"""
Initial Setup: Create named ranges and commission formula task
Task ID: calc_nrv_023
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_023'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Headers
    ws.cell(row=1, column=1, value="Salesperson")
    ws.cell(row=1, column=2, value="Sales")
    ws.cell(row=1, column=3, value="Commission")

    # G1 = 0.15 (commission rate value, no named range yet)
    ws.cell(row=1, column=7, value=0.15)

    # Salesperson data rows 2-20
    salespeople = [
        ("Sarah Chen", 45230),
        ("Marcus Johnson", 72150),
        ("Priya Patel", 38900),
        ("David Kim", 91200),
        ("Emma Rodriguez", 56700),
        ("James O'Brien", 64300),
        ("Aisha Mohammed", 83400),
        ("Carlos Rivera", 47800),
        ("Lisa Tanaka", 69500),
        ("Robert Williams", 52100),
        ("Maria Santos", 76800),
        ("Chen Wei", 41600),
        ("Anna Kowalski", 88300),
        ("Michael Brown", 35700),
        ("Fatima Al-Rashid", 97400),
        ("Thomas Lee", 43200),
        ("Jessica Park", 61800),
        ("Ahmed Hassan", 54900),
        ("Rachel Green", 79600),
    ]

    for i, (name, sales) in enumerate(salespeople, 2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=sales)
        # C column (Commission) left empty - task asks agent to write formula

    # Set column widths for readability
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['G'].width = 12

    # NO named ranges - the task asks the agent to create them
    # C2 is empty - the task asks the agent to write formula there

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
