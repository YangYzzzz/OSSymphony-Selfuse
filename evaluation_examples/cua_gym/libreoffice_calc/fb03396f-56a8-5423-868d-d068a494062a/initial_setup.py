"""
Initial Setup: Board games 2023 spreadsheet with empty Avg Play Time column.
Task ID: osworld_multi_apps_book_reading_rate_008
Domain: libreoffice_calc (multi-app: also creates longest_game.docx on Desktop)

Creates:
  - /home/user/boardgames_2023.xlsx  (open in LibreOffice Calc)
  - /home/user/Desktop/longest_game.docx  (empty .docx, open in LibreOffice Writer)
"""

import os
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, Alignment
from docx import Document

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_book_reading_rate_008'
SPREADSHEET = f'{WORKDIR}/boardgames_2023.xlsx'
DESKTOP_DOC = f'{WORKDIR}/Desktop/longest_game.docx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    # ----------------------------------------------------------------
    # 1. Create boardgames_2023.xlsx
    #    Columns: Game Name | Players | Avg Play Time (min)
    #    Avg Play Time intentionally EMPTY — agent must fill from BGG
    # ----------------------------------------------------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Board Games 2023"

    # Headers
    headers = ["Game Name", "Players", "Avg Play Time (min)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Data rows — Avg Play Time (col 3) is LEFT EMPTY intentionally
    games_data = [
        ("Twilight Imperium (4th Edition)", "3-6"),
        ("Gloomhaven",                      "1-4"),
        ("Wingspan",                         "1-5"),
        ("Pandemic Legacy Season 1",         "2-4"),
        ("Terraforming Mars",                "1-5"),
    ]
    for row_idx, (game, players) in enumerate(games_data, 2):
        ws.cell(row=row_idx, column=1, value=game)
        ws.cell(row=row_idx, column=2, value=players)
        # Column 3 (Avg Play Time) intentionally left empty

    # Column widths for readability
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 22

    wb.save(SPREADSHEET)
    print(f"Initial spreadsheet created: {SPREADSHEET}")

    # ----------------------------------------------------------------
    # 2. Create empty longest_game.docx on Desktop
    # ----------------------------------------------------------------
    os.makedirs(f"{WORKDIR}/Desktop", exist_ok=True)
    doc = Document()
    # Empty document — no paragraphs with text
    # (python-docx adds one empty paragraph by default, which is fine)
    doc.save(DESKTOP_DOC)
    print(f"Empty longest_game.docx created: {DESKTOP_DOC}")

    # ----------------------------------------------------------------
    # 3. GUI-ready startup: open both files
    # ----------------------------------------------------------------
    # Open spreadsheet in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{SPREADSHEET}"', delay_sec=2.0)
    # Open Writer document
    launch_gui(f'libreoffice --writer "{DESKTOP_DOC}"', delay_sec=2.0)
    print("GUI_READY: launched LibreOffice Calc and Writer with DISPLAY=:0")


create_initial()
