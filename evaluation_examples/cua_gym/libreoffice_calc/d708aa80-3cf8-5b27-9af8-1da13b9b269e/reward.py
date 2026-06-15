"""
FINAL REWARD SCRIPT - SUCCESS
Task: Summarize the total complaints for each issue type in Sheet2 with issue types as column headers using Pivot Table functionality.
Generated: 2025-11-24 07:47:59
Status: success
Model: o3
Total Steps: 2
"""

import openpyxl
import collections
import os

# -----------------------------------------------------------------------------
# Reward Script: Verify pivot-style summary of complaints by issue type
# -----------------------------------------------------------------------------
# Task to verify
# "Summarize the total complaints for each issue type in Sheet2 with issue types
#  as column headers using Pivot Table functionality."
#
# What we check:
# 1. Source sheet ("Complaints") still exists and contains an "Issue Type" column.
# 2. A secondary sheet (preferably named "Summary") exists.
# 3. The first non-empty row of that sheet contains the distinct issue types from
#    the source sheet – used here as column headers. (40 % of score)
# 4. The row immediately below those headers contains the corresponding totals
#    for each issue type.  (60 % of score)
#
# A perfect match of both headers and numbers yields REWARD 1.0.
# Partial credit is granted proportionally.
# -----------------------------------------------------------------------------

def verify_complaint_summary(file_path: str) -> float:
    """Return a score between 0.0 and 1.0 assessing task completion."""

    print(f"Starting verification for: {file_path}")
    total_score = 0.0  # progressive score
    MAX_SCORE = 1.0

    # ------------------------------------------------------------------
    # 1. Load workbook safely
    # ------------------------------------------------------------------
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        print("✓ Workbook loaded")
    except Exception as e:
        print(f"✗ Failed to load workbook: {e}")
        return 0.0

    # ------------------------------------------------------------------
    # 2. Verify source data sheet & column
    # ------------------------------------------------------------------
    if "Complaints" not in wb.sheetnames:
        print("✗ Sheet 'Complaints' missing – cannot verify task")
        return 0.0

    src = wb["Complaints"]
    issue_col_idx = None
    for cell in src[1]:
        if cell.value and str(cell.value).strip().lower() == "issue type":
            issue_col_idx = cell.column  # 1-based index
            break

    if issue_col_idx is None:
        print("✗ 'Issue Type' column not found in Complaints sheet")
        return 0.0

    # Gather expected counts from source data
    expected_counts = collections.Counter()
    for row in src.iter_rows(min_row=2, values_only=True):
        issue = row[issue_col_idx - 1]
        if issue is not None and str(issue).strip():
            expected_counts[str(issue).strip()] += 1

    print("Expected counts (from source):")
    for k, v in expected_counts.items():
        print(f"  {k}: {v}")

    # ------------------------------------------------------------------
    # 3. Locate summary sheet (any sheet other than 'Complaints')
    # ------------------------------------------------------------------
    summary = None
    if "Summary" in wb.sheetnames:
        summary = wb["Summary"]
    else:
        for name in wb.sheetnames:
            if name != "Complaints":
                summary = wb[name]
                break

    if summary is None:
        print("✗ No summary sheet found")
        return 0.0

    print(f"Using summary sheet: {summary.title}")

    # ------------------------------------------------------------------
    # 4. Extract header row (first non-empty row)
    # ------------------------------------------------------------------
    headers = None
    header_row_index = None
    for idx, row in enumerate(summary.iter_rows(values_only=True), start=1):
        if any(cell is not None for cell in row):
            headers = [str(c).strip() for c in row if c is not None]
            header_row_index = idx
            break

    if not headers:
        print("✗ No headers detected in summary sheet")
        return 0.0

    print("Headers found:", headers)

    # Score header coverage (40 %)
    correct_header_count = sum(1 for h in headers if h in expected_counts)
    header_coverage_ratio = correct_header_count / len(expected_counts) if expected_counts else 0
    header_score = 0.4 * header_coverage_ratio

    if header_score == 0:
        print("✗ Header names do not match any expected issue types")
    else:
        print(f"✓ Header coverage: {correct_header_count}/{len(expected_counts)} (score {header_score:.2f})")

    total_score += header_score

    # ------------------------------------------------------------------
    # 5. Retrieve counts row (row directly below headers)
    # ------------------------------------------------------------------
    counts_row_values = None
    if header_row_index is not None:
        next_rows = list(summary.iter_rows(min_row=header_row_index + 1,
                                            max_row=header_row_index + 1,
                                            values_only=True))
        if next_rows:
            counts_row_values = next_rows[0]

    if counts_row_values is None:
        print("✗ Could not find counts row under headers")
        # total_score already includes header portion; return it
        final = round(total_score, 2)
        print(f"Total score: {final}")
        return final

    print("Counts row values:", counts_row_values)

    # Map each header to its corresponding value
    header_to_value = {}
    for idx, header in enumerate(headers):
        if idx < len(counts_row_values):
            header_to_value[header] = counts_row_values[idx]

    # Score count accuracy (60 %)
    correct_counts = 0
    for issue, expected in expected_counts.items():
        reported = header_to_value.get(issue)
        if isinstance(reported, (int, float)) and abs(reported - expected) < 1e-6:
            correct_counts += 1
            print(f"  ✓ {issue}: expected {expected}, found {reported}")
        else:
            print(f"  ✗ {issue}: expected {expected}, found {reported}")

    count_accuracy_ratio = correct_counts / len(expected_counts) if expected_counts else 0
    count_score = 0.6 * count_accuracy_ratio
    total_score += count_score
    print(f"Count accuracy: {correct_counts}/{len(expected_counts)} (score {count_score:.2f})")

    # ------------------------------------------------------------------
    # 6. Finalize score (capped at 1.0)
    # ------------------------------------------------------------------
    final_score = round(min(total_score, MAX_SCORE), 2)
    print(f"Total score: {final_score}")
    return final_score


if __name__ == "__main__":
    # Path provided by the evaluation environment
    EXCEL_PATH = "/home/user/summarize_the_total_complaints_for_each_issue_type_in_sheet2_with_issue_types_as_column_headers_usin.xlsx"

    reward = verify_complaint_summary(EXCEL_PATH)
    print(f"REWARD: {reward}")
