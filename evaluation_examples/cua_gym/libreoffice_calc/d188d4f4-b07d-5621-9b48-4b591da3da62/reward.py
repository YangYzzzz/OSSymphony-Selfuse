"""
Reward Script: Set font color to red (#FF0000) for negative profit cells
Task ID: calc_fmt_font_color_negative_006
Domain: libreoffice_calc
Scoring:
  Component 1: All 4 target cells (D4, D7, D11, D15) have red font color FFFF0000 (0.6 pts)
  Component 2: Cell values in D4, D7, D11, D15 remain unchanged (-2340, -890, -4120, -670) (0.2 pts)
  Component 3: Non-target D column cells do NOT have red (#FF0000) font color applied (0.2 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmt_font_color_negative_006'


def get_font_color_rgb(cell):
    """
    Safely get a cell's font color RGB string.
    Returns the ARGB string (e.g., 'FFFF0000') if explicitly set,
    or None if using default/theme color (which throws ValueError).
    """
    try:
        rgb = cell.font.color.rgb
        return rgb
    except (ValueError, TypeError, AttributeError):
        # Default/theme color — no explicit RGB set
        return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Task: Manually set font color to red (#FF0000 / FFFF0000) for cells D4, D7, D11, D15
    which contain negative profit values (-2340, -890, -4120, -670).
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition gate: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: sheet 'Profit Analysis' must exist
    if 'Profit Analysis' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Profit Analysis' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Profit Analysis']

    # Component 1: All 4 target cells have red font color = FFFF0000 (0.6 pts)
    # The task asks to set font color to red (#FF0000) for D4, D7, D11, D15
    # In openpyxl ARGB format, #FF0000 stored as FFFF0000
    # This FAILS on initial (default/theme color, no explicit RGB) and PASSES on golden
    RED_ARGB = 'FFFF0000'
    target_cells = [(4, 4, 'D4'), (7, 4, 'D7'), (11, 4, 'D11'), (15, 4, 'D15')]

    try:
        red_cells_correct = []
        red_cells_wrong = []

        for row, col, coord in target_cells:
            cell = ws.cell(row=row, column=col)
            actual_color = get_font_color_rgb(cell)
            if actual_color == RED_ARGB:
                red_cells_correct.append(coord)
            else:
                red_cells_wrong.append(f"{coord}={repr(actual_color)}")

        n_correct = len(red_cells_correct)
        if n_correct == 4:
            print(f"PASS: Component 1 — All 4 target cells {[c for _, _, c in target_cells]} have red font color FFFF0000 (0.6 pts)")
            total_score += 0.6
        elif n_correct == 3:
            print(f"PARTIAL: Component 1 — 3/4 cells have red color. Correct: {red_cells_correct}, Wrong: {red_cells_wrong} (0.45 pts)")
            total_score += 0.45
        elif n_correct == 2:
            print(f"PARTIAL: Component 1 — 2/4 cells have red color. Correct: {red_cells_correct}, Wrong: {red_cells_wrong} (0.30 pts)")
            total_score += 0.30
        elif n_correct == 1:
            print(f"PARTIAL: Component 1 — 1/4 cells have red color. Correct: {red_cells_correct}, Wrong: {red_cells_wrong} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — No target cells have red font color. Wrong: {red_cells_wrong}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Cell values in D4, D7, D11, D15 remain unchanged (0.2 pts)
    # Ensures the formatting operation did not accidentally alter cell values.
    # Anchored to task completion: only awards points if Component 1 earned points
    # (i.e., some red coloring was applied), so initial file cannot earn this.
    expected_values = [
        (4, 4, 'D4', -2340),
        (7, 4, 'D7', -890),
        (11, 4, 'D11', -4120),
        (15, 4, 'D15', -670),
    ]
    try:
        value_issues = []
        for row, col, coord, expected_val in expected_values:
            actual_val = ws.cell(row=row, column=col).value
            if actual_val != expected_val:
                value_issues.append(f"{coord}: expected {expected_val}, found {actual_val}")

        # Gate: only award if some red coloring was applied (total_score > 0)
        if total_score > 0 and len(value_issues) == 0:
            print(f"PASS: Component 2 — All target cell values unchanged (-2340, -890, -4120, -670) (0.2 pts)")
            total_score += 0.2
        elif total_score > 0 and len(value_issues) > 0:
            print(f"FAIL: Component 2 — Some target cell values changed: {value_issues}")
        else:
            print(f"FAIL: Component 2 — Skipped (no red color applied in Component 1)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Non-target D column cells do NOT have red font color applied (0.2 pts)
    # Ensures only the 4 specified cells were colored, not all/extra D column cells.
    # Anchored to task completion: only awards points if Component 1 earned points,
    # so initial file cannot earn this component.
    non_target_rows = [1, 2, 3, 5, 6, 8, 9, 10, 12, 13, 14, 16, 17, 18, 19]
    try:
        non_target_red = []
        for row in non_target_rows:
            cell = ws.cell(row=row, column=4)
            actual_color = get_font_color_rgb(cell)
            if actual_color == RED_ARGB:
                non_target_red.append(f"D{row}(val={cell.value})")

        # Gate: only award if some red coloring was applied (total_score > 0)
        if total_score > 0 and len(non_target_red) == 0:
            print(f"PASS: Component 3 — Non-target D column cells have no red color applied (0.2 pts)")
            total_score += 0.2
        elif total_score > 0 and len(non_target_red) > 0:
            print(f"FAIL: Component 3 — Non-target cells incorrectly colored red: {non_target_red}")
        else:
            print(f"FAIL: Component 3 — Skipped (no red color applied in Component 1)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
