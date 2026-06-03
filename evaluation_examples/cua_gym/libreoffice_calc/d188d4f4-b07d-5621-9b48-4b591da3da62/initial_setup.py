"""
Initial Setup: Create spreadsheet with profit analysis data where negative values exist
Task ID: calc_fmt_font_color_negative_006
Domain: libreoffice_calc

Creates a 'Profit Analysis' sheet with product data.
Cells D4=-2340, D7=-890, D11=-4120, D15=-670 are negative profit values.
All column D cells have default black font color (no special coloring applied).
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_font_color_negative_006'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Profit Analysis ---
    ws = wb.active
    ws.title = 'Profit Analysis'

    # Headers
    headers = ['Product', 'Revenue', 'Cost', 'Profit']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Product data (rows 2-20)
    # D column (Profit) = Revenue - Cost
    # Required negatives: D4=-2340, D7=-890, D11=-4120, D15=-670
    data = [
        # row 2
        ('Wireless Headphones',   18500,  12300,  6200),
        # row 3
        ('Laptop Stand',          9800,   6100,   3700),
        # row 4  — negative profit required: D4 = -2340
        ('USB-C Hub',             7200,   9540,  -2340),
        # row 5
        ('Mechanical Keyboard',   22300,  15800,  6500),
        # row 6
        ('Monitor Arm',           14600,  10200,  4400),
        # row 7  — negative profit required: D7 = -890
        ('Webcam HD 1080p',       5100,   5990,   -890),
        # row 8
        ('Ergonomic Mouse',       11200,   7600,  3600),
        # row 9
        ('Desk Lamp LED',          8700,   5300,  3400),
        # row 10
        ('Cable Management Kit',   4300,   2800,  1500),
        # row 11 — negative profit required: D11 = -4120
        ('Portable SSD 1TB',      12400,  16520, -4120),
        # row 12
        ('Bluetooth Speaker',     16800,  11200,  5600),
        # row 13
        ('Smart Power Strip',      9200,   6100,  3100),
        # row 14
        ('Laptop Cooling Pad',     7600,   5400,  2200),
        # row 15 — negative profit required: D15 = -670
        ('HDMI Switch 4-Port',     4800,   5470,  -670),
        # row 16
        ('Wireless Charger Pad',  10300,   7100,  3200),
        # row 17
        ('Screen Cleaning Kit',    3100,   1800,  1300),
        # row 18
        ('USB Hub 7-Port',         8900,   5800,  3100),
        # row 19
        ('Desk Organizer Pro',     5600,   3900,  1700),
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # All cells use default font (no special color applied)
    # Column D cells intentionally have NO explicit font color set
    # so they render with the default black color in LibreOffice

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet: Profit Analysis, Rows 1-20')
    print(f'Negative profit cells: D4={ws["D4"].value}, D7={ws["D7"].value}, D11={ws["D11"].value}, D15={ws["D15"].value}')
    print('All column D cells have default black font color (no red applied).')


create_initial()
