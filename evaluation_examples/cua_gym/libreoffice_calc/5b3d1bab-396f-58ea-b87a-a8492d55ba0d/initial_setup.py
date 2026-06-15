"""
Initial Setup: Import duty calculation spreadsheet for international purchase orders
Task ID: calc_ops_logistics_import_duty_048
Domain: libreoffice_calc
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_logistics_import_duty_048'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: ImportPOs ---
    ws1 = wb.active
    ws1.title = 'ImportPOs'

    # Headers (A-J)
    headers = [
        'PO Number', 'HS Code', 'Description', 'Invoice Value USD',
        'Freight USD', 'Insurance USD', 'Duty Rate', 'Duty Amount',
        'Total Landed Cost', 'Landed Cost Markup %'
    ]
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # Realistic import PO data — 40 rows (rows 2-41)
    # Columns: PO Number, HS Code, Description, Invoice Value USD, Freight USD, Insurance USD
    # G, H, I, J left empty (that's what the agent needs to fill)
    po_data = [
        ('PO-2025-0001', '6109.10.00', 'Cotton T-Shirts',         18500.00,  920.00,  55.50),
        ('PO-2025-0002', '8471.30.01', 'Laptop Computers',        85000.00, 2100.00, 255.00),
        ('PO-2025-0003', '2106.90.99', 'Food Supplements',        12300.00,  615.00,  36.90),
        ('PO-2025-0004', '9401.61.00', 'Office Chairs',           22000.00, 1650.00,  66.00),
        ('PO-2025-0005', '3304.99.00', 'Cosmetic Creams',          9800.00,  490.00,  29.40),
        ('PO-2025-0006', '6109.10.00', 'Cotton Polo Shirts',      14200.00,  710.00,  42.60),
        ('PO-2025-0007', '8517.12.00', 'Mobile Phones',           73500.00, 1470.00, 220.50),
        ('PO-2025-0008', '9403.20.00', 'Steel Filing Cabinets',   31000.00, 2325.00,  93.00),
        ('PO-2025-0009', '2106.90.99', 'Protein Powders',         16800.00,  840.00,  50.40),
        ('PO-2025-0010', '8471.30.01', 'Desktop Computers',       62000.00, 1550.00, 186.00),
        ('PO-2025-0011', '3304.99.00', 'Facial Serums',           11400.00,  570.00,  34.20),
        ('PO-2025-0012', '6203.42.90', 'Denim Jeans',             27600.00, 1380.00,  82.80),
        ('PO-2025-0013', '8528.72.00', 'LCD Televisions',         58000.00, 2320.00, 174.00),
        ('PO-2025-0014', '9401.61.00', 'Conference Room Chairs',  44000.00, 3300.00, 132.00),
        ('PO-2025-0015', '2204.21.00', 'Red Wine Bottles',         8900.00,  801.00,  26.70),
        ('PO-2025-0016', '8517.12.00', 'Smartphones Premium',     92000.00, 1840.00, 276.00),
        ('PO-2025-0017', '6203.42.90', 'Men Cargo Pants',         19500.00,  975.00,  58.50),
        ('PO-2025-0018', '9403.20.00', 'Metal Shelving Units',    25400.00, 1905.00,  76.20),
        ('PO-2025-0019', '3304.99.00', 'Anti-Aging Moisturizer',  13600.00,  680.00,  40.80),
        ('PO-2025-0020', '2204.21.00', 'White Wine Assortment',    7200.00,  648.00,  21.60),
        ('PO-2025-0021', '6109.10.00', 'Cotton Sweatshirts',      21000.00, 1050.00,  63.00),
        ('PO-2025-0022', '8471.30.01', 'Workstation Computers',  105000.00, 2625.00, 315.00),
        ('PO-2025-0023', '2106.90.99', 'Vitamin Supplements',      9100.00,  455.00,  27.30),
        ('PO-2025-0024', '9401.61.00', 'Ergonomic Task Chairs',   38500.00, 2887.50, 115.50),
        ('PO-2025-0025', '3304.99.00', 'Sunscreen SPF50',          6700.00,  335.00,  20.10),
        ('PO-2025-0026', '8517.12.00', 'Rugged Smartphones',      48000.00,  960.00, 144.00),
        ('PO-2025-0027', '6203.42.90', 'Women Jeans Premium',     33200.00, 1660.00,  99.60),
        ('PO-2025-0028', '8528.72.00', 'LED Smart TVs 55 inch',   71000.00, 2840.00, 213.00),
        ('PO-2025-0029', '9403.20.00', 'Industrial Racking',      54000.00, 4050.00, 162.00),
        ('PO-2025-0030', '2204.21.00', 'Sparkling Wine',           5500.00,  495.00,  16.50),
        ('PO-2025-0031', '6109.10.00', 'Cotton Hoodies',          17800.00,  890.00,  53.40),
        ('PO-2025-0032', '2106.90.99', 'Omega-3 Capsules',        10600.00,  530.00,  31.80),
        ('PO-2025-0033', '8471.30.01', 'Server Equipment',       148000.00, 3700.00, 444.00),
        ('PO-2025-0034', '3304.99.00', 'Eye Cream Luxury',        15200.00,  760.00,  45.60),
        ('PO-2025-0035', '9401.61.00', 'Executive Office Chairs', 67500.00, 5062.50, 202.50),
        ('PO-2025-0036', '8517.12.00', 'Budget Smartphones',      29000.00,  580.00,  87.00),
        ('PO-2025-0037', '6203.42.90', 'Stretch Denim Jeans',     24800.00, 1240.00,  74.40),
        ('PO-2025-0038', '8528.72.00', '4K OLED Television',     115000.00, 4600.00, 345.00),
        ('PO-2025-0039', '9403.20.00', 'Modular Workbenches',     41000.00, 3075.00, 123.00),
        ('PO-2025-0040', '2204.21.00', 'Premium Red Wine Cases',  11200.00, 1008.00,  33.60),
    ]

    for r, row_data in enumerate(po_data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)
        # Columns G, H, I, J (7-10) left empty intentionally

    # --- Sheet 2: DutyRates ---
    ws2 = wb.create_sheet('DutyRates')

    # Headers
    dr_headers = ['HS Code', 'Product Category', 'Duty Rate %']
    for col, h in enumerate(dr_headers, 1):
        ws2.cell(row=1, column=col, value=h)

    # 20 HS code duty rate entries
    duty_rates = [
        ('6109.10.00', 'Cotton Knit Tops (T-Shirts, Polos, Sweatshirts)', 0.12),
        ('6203.42.90', 'Men/Women Woven Trousers & Jeans',                0.10),
        ('8471.30.01', 'Computers (Laptops, Desktops, Workstations)',     0.00),
        ('8517.12.00', 'Mobile/Smartphones & Cellular Devices',           0.00),
        ('8528.72.00', 'Color Television Receivers (LCD/LED)',            0.05),
        ('9401.61.00', 'Upholstered Seats (Office & Home Chairs)',        0.08),
        ('9403.20.00', 'Metal Furniture (Cabinets, Shelving, Racking)',   0.06),
        ('2204.21.00', 'Wine in Containers <= 2L',                        0.15),
        ('2106.90.99', 'Food Preparations NEC (Supplements)',             0.07),
        ('3304.99.00', 'Cosmetics & Skin Care Preparations',              0.09),
        ('8443.31.00', 'Inkjet Printing Machines',                        0.00),
        ('8703.22.10', 'Passenger Vehicles 1000-1500cc',                  0.25),
        ('6204.62.90', 'Women Woven Cotton Trousers',                     0.10),
        ('8507.60.00', 'Lithium-Ion Batteries',                           0.03),
        ('9506.62.00', 'Sports Balls & Inflatable Equipment',             0.05),
        ('8516.60.00', 'Ovens, Cookers, Kitchen Appliances',              0.08),
        ('9608.10.10', 'Ball Point Pens',                                 0.06),
        ('4202.12.00', 'Trunks, Suitcases, Travel Bags',                 0.12),
        ('8544.42.90', 'Electric Conductors/Cables',                      0.04),
        ('7117.19.00', 'Imitation Jewellery Base Metal',                  0.15),
    ]

    for r, row_data in enumerate(duty_rates, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  ImportPOs sheet: 40 data rows, G-J columns empty')
    print(f'  DutyRates sheet: 20 duty rate entries')


create_initial()
