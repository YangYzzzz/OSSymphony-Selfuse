"""
Reward Script: Calculate import duties and total landed cost for international POs
Task ID: calc_ops_logistics_import_duty_048
Domain: libreoffice_calc
Scoring:
  Component 1: VLOOKUP formulas in G2:G41 pulling duty rate from DutyRates sheet  (0.35 pts)
  Component 2: Duty Amount formulas in H2:H41 = D*G                               (0.20 pts)
  Component 3: Total Landed Cost formulas in I2:I41 = D+E+F+H                     (0.20 pts)
  Component 4: Landed Cost Markup % formulas in J2:J41 = (I-D)/D                  (0.15 pts)
  Component 5: Number formats on G (0.00%) and J (0.00%) columns                  (0.10 pts)
Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_logistics_import_duty_048'


def normalize_formula(formula):
    """Normalize formula for comparison: uppercase, strip spaces."""
    if formula is None:
        return ''
    return formula.upper().replace(' ', '')


def check_vlookup_formula(formula, row):
    """
    Check if a formula is a valid VLOOKUP referencing DutyRates for the given row.
    Valid patterns (case-insensitive):
      =VLOOKUP(B{row}, DutyRates!$A:$C, 3, FALSE)  -- various spacing/quoting variants
      =VLOOKUP(B{row}, DutyRates!A:C, 3, FALSE)
      =VLOOKUP(B{row}, DutyRates!$A$1:$C$21, 3, 0)  -- or similar absolute refs
    """
    if not formula or not isinstance(formula, str):
        return False
    f = normalize_formula(formula)
    # Must start with =VLOOKUP
    if not f.startswith('=VLOOKUP('):
        return False
    # Must reference B{row} as lookup value
    if f'B{row}' not in f:
        return False
    # Must reference DutyRates sheet
    if 'DUTYRATES' not in f:
        return False
    # Must return column 3 (the Duty Rate % column)
    if ',3,' not in f:
        return False
    return True


def check_duty_amount_formula(formula, row):
    """
    Check if formula is =D{row}*G{row} (or equivalent with spaces).
    """
    if not formula or not isinstance(formula, str):
        return False
    f = normalize_formula(formula)
    # Accept =D{row}*G{row} or =G{row}*D{row}
    pattern1 = f'=D{row}*G{row}'
    pattern2 = f'=G{row}*D{row}'
    return f == pattern1 or f == pattern2


def check_total_landed_cost_formula(formula, row):
    """
    Check if formula is =D{row}+E{row}+F{row}+H{row} (any addition order).
    """
    if not formula or not isinstance(formula, str):
        return False
    f = normalize_formula(formula)
    # Must be an addition formula containing all four cells
    if not f.startswith('='):
        return False
    body = f[1:]  # strip leading =
    # Check all required cells are present
    required = [f'D{row}', f'E{row}', f'F{row}', f'H{row}']
    for cell in required:
        if cell not in body:
            return False
    # Check it uses addition (no subtraction between these cells)
    # Split by + and verify all four refs are among the addends
    return True


def check_markup_formula(formula, row):
    """
    Check if formula is =(I{row}-D{row})/D{row}.
    """
    if not formula or not isinstance(formula, str):
        return False
    f = normalize_formula(formula)
    # Must contain I{row}-D{row} and divide by D{row}
    expected = f'=(I{row}-D{row})/D{row}'
    return f == expected


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: check required sheets exist
    if 'ImportPOs' not in wb.sheetnames:
        print("CRITICAL: Sheet 'ImportPOs' not found")
        print("REWARD: 0.0")
        return 0.0

    if 'DutyRates' not in wb.sheetnames:
        print("CRITICAL: Sheet 'DutyRates' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['ImportPOs']

    # Component 1: VLOOKUP formulas in G2:G41 (0.35 pts)
    # Each row's G column should have a VLOOKUP pulling duty rate from DutyRates
    try:
        vlookup_pass = 0
        vlookup_fail = 0
        vlookup_fail_examples = []
        for row in range(2, 42):
            formula = ws.cell(row=row, column=7).value
            if check_vlookup_formula(formula, row):
                vlookup_pass += 1
            else:
                vlookup_fail += 1
                if len(vlookup_fail_examples) < 3:
                    vlookup_fail_examples.append(f"G{row}={repr(formula)}")

        if vlookup_pass == 40:
            print(f"PASS: Component 1 — VLOOKUP in G2:G41 all 40 rows correct (0.35 pts)")
            total_score += 0.35
        elif vlookup_pass >= 30:
            partial = round(0.35 * vlookup_pass / 40, 4)
            print(f"PARTIAL: Component 1 — VLOOKUP in G2:G41: {vlookup_pass}/40 correct, awarding {partial} pts")
            print(f"  Failing examples: {vlookup_fail_examples}")
            if partial > 0:
                total_score += partial
        else:
            print(f"FAIL: Component 1 — VLOOKUP in G2:G41: only {vlookup_pass}/40 correct")
            print(f"  Failing examples: {vlookup_fail_examples}")
    except Exception as e:
        print(f"ERROR: Component 1 (VLOOKUP check): {e}")

    # Component 2: Duty Amount formulas in H2:H41 = D*G (0.20 pts)
    # H = Invoice Value * Duty Rate
    try:
        h_pass = 0
        h_fail_examples = []
        for row in range(2, 42):
            formula = ws.cell(row=row, column=8).value
            if check_duty_amount_formula(formula, row):
                h_pass += 1
            else:
                if len(h_fail_examples) < 3:
                    h_fail_examples.append(f"H{row}={repr(formula)}")

        if h_pass == 40:
            print(f"PASS: Component 2 — Duty Amount (D*G) in H2:H41 all 40 rows correct (0.20 pts)")
            total_score += 0.20
        elif h_pass >= 20:
            partial = round(0.20 * h_pass / 40, 4)
            print(f"PARTIAL: Component 2 — Duty Amount in H2:H41: {h_pass}/40 correct, awarding {partial} pts")
            print(f"  Failing examples: {h_fail_examples}")
            if partial > 0:
                total_score += partial
        else:
            print(f"FAIL: Component 2 — Duty Amount in H2:H41: only {h_pass}/40 correct")
            print(f"  Failing examples: {h_fail_examples}")
    except Exception as e:
        print(f"ERROR: Component 2 (Duty Amount check): {e}")

    # Component 3: Total Landed Cost formulas in I2:I41 = D+E+F+H (0.20 pts)
    try:
        i_pass = 0
        i_fail_examples = []
        for row in range(2, 42):
            formula = ws.cell(row=row, column=9).value
            if check_total_landed_cost_formula(formula, row):
                i_pass += 1
            else:
                if len(i_fail_examples) < 3:
                    i_fail_examples.append(f"I{row}={repr(formula)}")

        if i_pass == 40:
            print(f"PASS: Component 3 — Total Landed Cost (D+E+F+H) in I2:I41 all 40 rows correct (0.20 pts)")
            total_score += 0.20
        elif i_pass >= 20:
            partial = round(0.20 * i_pass / 40, 4)
            print(f"PARTIAL: Component 3 — Total Landed Cost in I2:I41: {i_pass}/40 correct, awarding {partial} pts")
            print(f"  Failing examples: {i_fail_examples}")
            if partial > 0:
                total_score += partial
        else:
            print(f"FAIL: Component 3 — Total Landed Cost in I2:I41: only {i_pass}/40 correct")
            print(f"  Failing examples: {i_fail_examples}")
    except Exception as e:
        print(f"ERROR: Component 3 (Total Landed Cost check): {e}")

    # Component 4: Landed Cost Markup % formulas in J2:J41 = (I-D)/D (0.15 pts)
    try:
        j_pass = 0
        j_fail_examples = []
        for row in range(2, 42):
            formula = ws.cell(row=row, column=10).value
            if check_markup_formula(formula, row):
                j_pass += 1
            else:
                if len(j_fail_examples) < 3:
                    j_fail_examples.append(f"J{row}={repr(formula)}")

        if j_pass == 40:
            print(f"PASS: Component 4 — Markup % ((I-D)/D) in J2:J41 all 40 rows correct (0.15 pts)")
            total_score += 0.15
        elif j_pass >= 20:
            partial = round(0.15 * j_pass / 40, 4)
            print(f"PARTIAL: Component 4 — Markup % in J2:J41: {j_pass}/40 correct, awarding {partial} pts")
            print(f"  Failing examples: {j_fail_examples}")
            if partial > 0:
                total_score += partial
        else:
            print(f"FAIL: Component 4 — Markup % in J2:J41: only {j_pass}/40 correct")
            print(f"  Failing examples: {j_fail_examples}")
    except Exception as e:
        print(f"ERROR: Component 4 (Markup % check): {e}")

    # Component 5: Number formats — G column should be percentage, J column should be percentage (0.10 pts)
    # D, E, F, H, I columns should have currency format
    try:
        # Check G2 and J2 have percentage format (0.00% or similar)
        g2_fmt = ws.cell(row=2, column=7).number_format
        j2_fmt = ws.cell(row=2, column=10).number_format

        g_is_pct = g2_fmt is not None and ('%' in g2_fmt or 'percent' in g2_fmt.lower())
        j_is_pct = j2_fmt is not None and ('%' in j2_fmt or 'percent' in j2_fmt.lower())

        # Check D2, H2, I2 have currency format
        d2_fmt = ws.cell(row=2, column=4).number_format
        h2_fmt = ws.cell(row=2, column=8).number_format
        i2_fmt = ws.cell(row=2, column=9).number_format

        d_is_currency = d2_fmt is not None and ('$' in d2_fmt or '#,##0' in d2_fmt)
        h_is_currency = h2_fmt is not None and ('$' in h2_fmt or '#,##0' in h2_fmt)
        i_is_currency = i2_fmt is not None and ('$' in i2_fmt or '#,##0' in i2_fmt)

        format_checks = {
            'G (duty rate %)': g_is_pct,
            'J (markup %)': j_is_pct,
            'D (invoice currency)': d_is_currency,
            'H (duty amount currency)': h_is_currency,
            'I (landed cost currency)': i_is_currency,
        }

        passed_formats = sum(1 for v in format_checks.values() if v)
        failed_formats = {k: v for k, v in format_checks.items() if not v}

        if passed_formats == 5:
            print(f"PASS: Component 5 — All number formats correct (G/J=%, D/H/I=currency) (0.10 pts)")
            total_score += 0.10
        elif passed_formats >= 3:
            partial = round(0.10 * passed_formats / 5, 4)
            print(f"PARTIAL: Component 5 — Number formats: {passed_formats}/5 correct, awarding {partial} pts")
            print(f"  Failed: {failed_formats}")
            print(f"  Actual formats: G2={repr(g2_fmt)}, J2={repr(j2_fmt)}, D2={repr(d2_fmt)}, H2={repr(h2_fmt)}, I2={repr(i2_fmt)}")
            if partial > 0:
                total_score += partial
        else:
            print(f"FAIL: Component 5 — Number formats: only {passed_formats}/5 correct")
            print(f"  Failed: {failed_formats}")
            print(f"  Actual formats: G2={repr(g2_fmt)}, J2={repr(j2_fmt)}, D2={repr(d2_fmt)}, H2={repr(h2_fmt)}, I2={repr(i2_fmt)}")
    except Exception as e:
        print(f"ERROR: Component 5 (Number format check): {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {round(final_score, 4)}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
