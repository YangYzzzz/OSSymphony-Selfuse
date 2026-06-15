"""
Initial Setup: Survey Responses spreadsheet with inconsistent satisfaction level text values.
Task ID: calc_dop_validate_dropdown_055
Domain: libreoffice_calc
"""

import os
import random
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_dop_validate_dropdown_055'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

random.seed(42)

AGE_GROUPS = ['18-24', '25-34', '35-44', '45-54', '55-64', '65+']
REGIONS = ['North America', 'Europe', 'Asia Pacific', 'Latin America', 'Middle East', 'Africa']

# Inconsistent free-text values for Satisfaction Level (before validation)
INCONSISTENT_SATISFACTION = [
    'great', 'ok', 'terrible', 'awful', 'good', 'bad', 'fine', 'excellent',
    'poor', 'very good', 'not bad', 'outstanding', 'horrible', 'decent',
    'mediocre', 'amazing', 'disappointing', 'satisfactory', 'unsatisfactory',
    'fantastic', 'okay', 'Not Happy', 'loved it', 'hate it', 'meh', 'alright',
    'so so', '5 stars', '1 star', 'best ever', 'worst ever', 'average',
    'above average', 'below average', 'top notch', 'needs improvement'
]

FIRST_NAMES = [
    'James', 'Mary', 'John', 'Patricia', 'Robert', 'Jennifer', 'Michael', 'Linda',
    'William', 'Barbara', 'David', 'Elizabeth', 'Richard', 'Susan', 'Joseph', 'Jessica',
    'Thomas', 'Sarah', 'Charles', 'Karen', 'Christopher', 'Lisa', 'Daniel', 'Nancy',
    'Matthew', 'Betty', 'Anthony', 'Margaret', 'Mark', 'Sandra', 'Donald', 'Ashley',
    'Steven', 'Dorothy', 'Paul', 'Kimberly', 'Andrew', 'Emily', 'Kenneth', 'Donna',
    'George', 'Michelle', 'Joshua', 'Carol', 'Kevin', 'Amanda', 'Brian', 'Melissa',
    'Edward', 'Deborah', 'Ronald', 'Stephanie', 'Timothy', 'Rebecca', 'Jason', 'Sharon',
    'Jeffrey', 'Laura', 'Ryan', 'Cynthia', 'Jacob', 'Kathleen', 'Gary', 'Amy'
]

LAST_NAMES = [
    'Smith', 'Johnson', 'Williams', 'Brown', 'Jones', 'Garcia', 'Miller', 'Davis',
    'Rodriguez', 'Martinez', 'Hernandez', 'Lopez', 'Gonzalez', 'Wilson', 'Anderson',
    'Thomas', 'Taylor', 'Moore', 'Jackson', 'Martin', 'Lee', 'Perez', 'Thompson',
    'White', 'Harris', 'Sanchez', 'Clark', 'Ramirez', 'Lewis', 'Robinson', 'Walker',
    'Young', 'Allen', 'King', 'Wright', 'Scott', 'Torres', 'Nguyen', 'Hill', 'Flores',
    'Green', 'Adams', 'Nelson', 'Baker', 'Hall', 'Rivera', 'Campbell', 'Mitchell',
    'Carter', 'Roberts', 'Gomez', 'Phillips', 'Evans', 'Turner', 'Diaz', 'Parker',
    'Cruz', 'Edwards', 'Collins', 'Reyes', 'Stewart', 'Morris', 'Morales', 'Murphy'
]

COMMENT_TEMPLATES = [
    'Overall good experience with the service.',
    'Response time could be improved significantly.',
    'The support team was very helpful and professional.',
    'I had some technical difficulties during the process.',
    'Would recommend to friends and colleagues.',
    'The interface is intuitive and easy to use.',
    'Some features are missing that would be useful.',
    'Excellent value for money.',
    'Had to wait too long for a resolution.',
    'The documentation needs to be clearer.',
    'Very impressed with the quality of service.',
    'Issues were resolved quickly and efficiently.',
    'Could use more customization options.',
    'The team went above and beyond to help.',
    'Performance was slow during peak hours.',
    'Easy to navigate and find what I needed.',
    'Communication was clear throughout the process.',
    'Pricing seems a bit high for what you get.',
    'Integration with other tools works seamlessly.',
    'Had to contact support multiple times for one issue.',
    'Setup was straightforward and took minimal time.',
    'Mobile experience needs improvement.',
    'Data export features are very useful.',
    'Would like more reporting capabilities.',
    'Security features give me peace of mind.',
]

def generate_timestamp(row_num):
    """Generate a realistic timestamp for a survey response."""
    year = 2024
    month = random.randint(1, 12)
    day = random.randint(1, 28)
    hour = random.randint(8, 20)
    minute = random.randint(0, 59)
    return f'{year}-{month:02d}-{day:02d} {hour:02d}:{minute:02d}'

def generate_email(first, last, row_num):
    """Generate a plausible email address."""
    domains = ['gmail.com', 'yahoo.com', 'outlook.com', 'hotmail.com', 'company.com', 'work.org']
    domain = domains[row_num % len(domains)]
    return f'{first.lower()}.{last.lower()}{row_num % 100}@{domain}'

def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: SurveyResponses ---
    ws = wb.active
    ws.title = 'SurveyResponses'

    # Headers
    headers = ['Response ID', 'Timestamp', 'Email', 'Age Group', 'Region', 'Satisfaction Level', 'Comments']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Data rows (499 rows, rows 2 through 500)
    for i in range(499):
        row_num = i + 2
        resp_id = f'RESP-{2024 * 100 + i + 1:07d}'
        first = random.choice(FIRST_NAMES)
        last = random.choice(LAST_NAMES)
        timestamp = generate_timestamp(i)
        email = generate_email(first, last, i + 1)
        age_group = random.choice(AGE_GROUPS)
        region = random.choice(REGIONS)
        # Inconsistent free-text satisfaction level (task requirement: column F has no validation)
        satisfaction = random.choice(INCONSISTENT_SATISFACTION)
        comment = random.choice(COMMENT_TEMPLATES)

        ws.cell(row=row_num, column=1, value=resp_id)
        ws.cell(row=row_num, column=2, value=timestamp)
        ws.cell(row=row_num, column=3, value=email)
        ws.cell(row=row_num, column=4, value=age_group)
        ws.cell(row=row_num, column=5, value=region)
        ws.cell(row=row_num, column=6, value=satisfaction)
        ws.cell(row=row_num, column=7, value=comment)

    # Set some column widths for readability
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 45

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet: SurveyResponses')
    print(f'Rows: 499 data rows (rows 2-500)')
    print(f'Column F: inconsistent free-text, NO data validation')

create_initial()
