"""
Initial Setup: Create pivot-table-style report with raw transaction data and empty Pivot sheet
Task ID: calc_sales_058
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_058'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Transactions ---
    ws1 = wb.active
    ws1.title = 'Transactions'

    # Headers
    headers = ['Region', 'Product', 'Quarter', 'Revenue']
    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font

    # Transaction data (8 rows as specified in context)
    data = [
        ['North', 'SaaS',  'Q1', 40000],
        ['North', 'HW',    'Q1', 25000],
        ['South', 'SaaS',  'Q1', 35000],
        ['North', 'SaaS',  'Q2', 55000],
        ['South', 'HW',    'Q2', 30000],
        ['South', 'SaaS',  'Q2', 42000],
        ['North', 'HW',    'Q2', 38000],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Set reasonable column widths
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 14

    # Format revenue column as currency
    for r in range(2, 9):
        ws1.cell(row=r, column=4).number_format = '#,##0'

    # --- Sheet 2: Pivot (labels only, NO formulas or values in B2:D6) ---
    ws2 = wb.create_sheet('Pivot')

    # Header row: A1 blank, B1='Q1', C1='Q2', D1='Total'
    ws2.cell(row=1, column=2, value='Q1')
    ws2.cell(row=1, column=3, value='Q2')
    ws2.cell(row=1, column=4, value='Total')

    # Style headers
    for col in range(2, 5):
        cell = ws2.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Row labels: A2:A6
    row_labels = ['North-SaaS', 'North-HW', 'South-SaaS', 'South-HW', 'Grand Total']
    for r, label in enumerate(row_labels, 2):
        cell = ws2.cell(row=r, column=1, value=label)
        if label == 'Grand Total':
            cell.font = Font(bold=True)

    # Set column widths
    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 12

    # B2:D6 are intentionally LEFT EMPTY - the task is to fill them with SUMIFS formulas

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
