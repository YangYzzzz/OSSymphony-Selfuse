"""
Reward Script: Pivot-table-style report with SUMIFS, row totals, and grand total
Task ID: calc_sales_058
Domain: libreoffice_calc
Scoring:
  Component 1 (0.50): SUMIFS formulas in B2:C5 referencing correct region/product/quarter (8 cells)
  Component 2 (0.25): Row total formulas in D2:D5
  Component 3 (0.25): Grand Total formulas in B6, C6, D6
All components check task-introduced changes only (formulas in Pivot sheet).
Initial env has empty Pivot cells -> scores 0.0. Golden env has formulas -> scores 1.0.
"""

import os

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_058'

# Expected region/product combos for each row
ROW_PARAMS = {
    2: ('North', 'SaaS'),
    3: ('North', 'HW'),
    4: ('South', 'SaaS'),
    5: ('South', 'HW'),
}


def is_sumifs_formula(val, region, product, quarter):
    """Check if a cell value is a SUMIFS formula referencing the correct parameters."""
    if not isinstance(val, str):
        return False
    v = val.upper().replace(' ', '')
    if not v.startswith('=SUMIFS('):
        return False
    if f'"{region.upper()}"' not in v:
        return False
    if f'"{product.upper()}"' not in v:
        return False
    if f'"{quarter.upper()}"' not in v:
        return False
    # Must reference the Transactions sheet data
    if 'TRANSACTIONS!' not in v:
        return False
    return True


def is_row_total_formula(val, row):
    """Check if a cell value is a sum/addition formula for B+C in the same row."""
    if not isinstance(val, str):
        return False
    v = val.upper().replace(' ', '')
    if not v.startswith('='):
        return False
    # Accept =B{row}+C{row} or =SUM(B{row}:C{row}) or similar
    if f'B{row}' in v and f'C{row}' in v:
        return True
    return False


def is_column_sum_formula(val, col, start_row=2, end_row=5):
    """Check if a cell value is a SUM formula for a column range."""
    if not isinstance(val, str):
        return False
    v = val.upper().replace(' ', '')
    if not v.startswith('='):
        return False
    # Accept =SUM(X2:X5) or similar patterns
    if 'SUM(' in v and f'{col}{start_row}' in v and f'{col}{end_row}' in v:
        return True
    # Also accept explicit addition: =X2+X3+X4+X5
    refs = [f'{col}{r}' for r in range(start_row, end_row + 1)]
    if all(ref in v for ref in refs):
        return True
    return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Pivot sheet must exist
    if 'Pivot' not in wb.sheetnames:
        print("FAIL: 'Pivot' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Pivot']

    # Component 1: SUMIFS formulas in B2:C5 (0.50 points)
    # Each of the 8 cells gets 0.0625 points
    try:
        sumifs_count = 0
        for row_num, (region, product) in ROW_PARAMS.items():
            for col, quarter in [('B', 'Q1'), ('C', 'Q2')]:
                cell_ref = f'{col}{row_num}'
                val = ws[cell_ref].value
                if is_sumifs_formula(val, region, product, quarter):
                    sumifs_count += 1
                else:
                    print(f"FAIL: {cell_ref} — expected SUMIFS for {region}/{product}/{quarter}, found: {repr(val)}")

        comp1_score = (sumifs_count / 8) * 0.50
        if sumifs_count == 8:
            print(f"PASS: Component 1 — All 8 SUMIFS formulas correct ({comp1_score:.2f} pts)")
            total_score += comp1_score
        elif sumifs_count > 0:
            print(f"PARTIAL: Component 1 — {sumifs_count}/8 SUMIFS formulas correct ({comp1_score:.4f} pts)")
            total_score += comp1_score
        else:
            print(f"FAIL: Component 1 — No valid SUMIFS formulas found (0.00 pts)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Row total formulas in D2:D5 (0.25 points)
    # Each of 4 cells gets 0.0625 points
    try:
        row_total_count = 0
        for row_num in range(2, 6):
            cell_ref = f'D{row_num}'
            val = ws[cell_ref].value
            if is_row_total_formula(val, row_num):
                row_total_count += 1
            else:
                print(f"FAIL: {cell_ref} — expected row total formula, found: {repr(val)}")

        comp2_score = (row_total_count / 4) * 0.25
        if row_total_count == 4:
            print(f"PASS: Component 2 — All 4 row total formulas correct ({comp2_score:.2f} pts)")
            total_score += comp2_score
        elif row_total_count > 0:
            print(f"PARTIAL: Component 2 — {row_total_count}/4 row total formulas correct ({comp2_score:.4f} pts)")
            total_score += comp2_score
        else:
            print(f"FAIL: Component 2 — No valid row total formulas found (0.00 pts)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Grand Total formulas in B6, C6, D6 (0.25 points)
    # Each of 3 cells gets ~0.0833 points
    try:
        gt_count = 0
        for col in ['B', 'C']:
            cell_ref = f'{col}6'
            val = ws[cell_ref].value
            if is_column_sum_formula(val, col):
                gt_count += 1
            else:
                print(f"FAIL: {cell_ref} — expected column SUM formula, found: {repr(val)}")

        # D6 can be SUM(D2:D5), SUM(B6:C6), B6+C6, or similar
        d6_val = ws['D6'].value
        if isinstance(d6_val, str):
            d6_upper = d6_val.upper().replace(' ', '')
            if d6_upper.startswith('=') and (
                is_column_sum_formula(d6_val, 'D') or
                ('B6' in d6_upper and 'C6' in d6_upper) or
                ('D2' in d6_upper and 'D5' in d6_upper)
            ):
                gt_count += 1
            else:
                print(f"FAIL: D6 — expected grand total formula, found: {repr(d6_val)}")
        else:
            print(f"FAIL: D6 — expected formula, found: {repr(d6_val)}")

        comp3_score = (gt_count / 3) * 0.25
        if gt_count == 3:
            print(f"PASS: Component 3 — All 3 grand total formulas correct ({comp3_score:.2f} pts)")
            total_score += comp3_score
        elif gt_count > 0:
            print(f"PARTIAL: Component 3 — {gt_count}/3 grand total formulas correct ({comp3_score:.4f} pts)")
            total_score += comp3_score
        else:
            print(f"FAIL: Component 3 — No valid grand total formulas found (0.00 pts)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score:.4f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved LibreOffice state before verification
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
