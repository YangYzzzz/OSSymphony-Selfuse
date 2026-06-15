"""
Initial Setup: Daily Sales 2023 spreadsheet for SUMIFS date range task
Task ID: calc_fmb_sumif_date_range_051
Domain: libreoffice_calc
"""

import os
import random
from datetime import date, timedelta
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_sumif_date_range_051'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Daily Sales 2023'

    # --- Row 1: Headers ---
    ws['A1'] = 'Date'
    ws['B1'] = 'Sales Amount'
    ws['C1'] = 'Region'
    ws['D1'] = 'Rep'

    # --- Regions and Reps ---
    regions = ['North', 'South', 'East', 'West', 'Central']
    reps = [
        'Sarah Chen', 'Marcus Johnson', 'Priya Patel', 'David Kim',
        'Luisa Torres', 'James Carter', 'Emily Nguyen', 'Robert Walsh',
        'Amara Osei', 'Daniel Lee'
    ]

    # --- Generate 364 days of 2023 (Jan 1 to Dec 30) ---
    # We need Q4 (Oct 1 - Dec 31) total = 1,234,780 exactly
    # Q4 days: Oct (31) + Nov (30) + Dec (31) = 92 days
    # Non-Q4 days: 364 - 92 = 272 days

    start_date = date(2023, 1, 1)
    q4_start = date(2023, 10, 1)
    q4_end = date(2023, 12, 31)

    all_dates = [start_date + timedelta(days=i) for i in range(364)]

    q4_dates = [d for d in all_dates if q4_start <= d <= q4_end]
    non_q4_dates = [d for d in all_dates if not (q4_start <= d <= q4_end)]

    # Generate non-Q4 sales: realistic daily amounts between 2000 and 8000
    non_q4_sales = []
    for _ in non_q4_dates:
        amount = round(random.uniform(2000, 8000), 2)
        non_q4_sales.append(amount)

    # Generate Q4 sales that sum to exactly 1,234,780
    # First generate random amounts for all but the last Q4 day
    q4_sales = []
    q4_target = 1234780.00
    running_total = 0.0

    for i in range(len(q4_dates) - 1):
        remaining_days = len(q4_dates) - i
        remaining_budget = q4_target - running_total
        # Keep amounts realistic while ensuring we can hit the target
        min_val = max(2000, remaining_budget / remaining_days - 5000)
        max_val = min(20000, remaining_budget / remaining_days + 5000)
        amount = round(random.uniform(min_val, max_val), 2)
        q4_sales.append(amount)
        running_total += amount

    # Last Q4 day gets the remainder
    last_q4 = round(q4_target - running_total, 2)
    q4_sales.append(last_q4)

    # Verify Q4 total
    assert abs(sum(q4_sales) - q4_target) < 0.01, f"Q4 total mismatch: {sum(q4_sales)}"

    # Build combined data list sorted by date
    date_sales_map = {}
    for d, s in zip(non_q4_dates, non_q4_sales):
        date_sales_map[d] = s
    for d, s in zip(q4_dates, q4_sales):
        date_sales_map[d] = s

    # Write data rows
    row = 2
    for d in all_dates:
        region = regions[(row - 2) % len(regions)]
        rep = reps[(row - 2) % len(reps)]
        ws.cell(row=row, column=1, value=d)
        ws.cell(row=row, column=2, value=date_sales_map[d])
        ws.cell(row=row, column=3, value=region)
        ws.cell(row=row, column=4, value=rep)
        row += 1

    # Row 365 is the last row (row 2 + 363 = row 365)
    # D2 should have label 'Q4 2023 Total', E2 should be empty (target)
    # Note: D2 is already written above as a rep name, we need to override
    # The context says D2 contains label 'Q4 2023 Total'
    # So D2 is a special label cell; we override it
    ws['D2'] = 'Q4 2023 Total'
    # E2 must be empty (target cell)
    ws['E2'] = None

    # Format date column as date
    from openpyxl.styles import numbers
    for r in range(2, 366):
        ws.cell(row=r, column=1).number_format = 'DD/MM/YYYY'

    # Column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Daily Sales 2023')
    print(f'  Rows: 1 header + 364 data rows')
    print(f'  Q4 2023 total in data: {sum(q4_sales):.2f}')
    print(f'  D2: Q4 2023 Total (label)')
    print(f'  E2: empty (target cell)')

create_initial()
