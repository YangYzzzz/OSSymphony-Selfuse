"""
Reward Script: Calculate total sales in Q4 2023 using SUMIFS formula
Task ID: calc_fmb_sumif_date_range_051
Domain: libreoffice_calc
Scoring:
  - Component 1: E2 contains a SUMIFS formula referencing correct data range (0.4 pts)
  - Component 2: E2 formula uses correct Q4 2023 date bounds (0.3 pts)
  - Component 3: Computed Q4 2023 total matches expected value ~1,234,780 (0.3 pts)
"""

import os
import re
import datetime
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmb_sumif_date_range_051'

def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    The task: Put a SUMIFS formula in cell E2 of 'Daily Sales 2023' sheet
    that calculates the total sales for Q4 2023 (Oct 1 - Dec 31, 2023).
    The formula should reference B2:B365 for amounts and A2:A365 for dates.
    Expected result: 1,234,780
    """
    total_score = 0.0

    # Load workbook (formula mode — to read the formula string)
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the expected sheet exists as a precondition gate
    if 'Daily Sales 2023' not in wb.sheetnames:
        print("FAIL: Sheet 'Daily Sales 2023' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Daily Sales 2023']

    # Component 1: E2 contains a SUMIFS formula referencing correct data ranges (0.4 pts)
    # The formula must use SUMIFS and reference B2:B365 for amounts and A2:A365 for dates.
    # This FAILS on initial (E2 is None) and PASSES on golden.
    try:
        e2_value = ws['E2'].value
        if e2_value is None or not isinstance(e2_value, str):
            print(f"FAIL: Component 1 — E2 is empty or not a formula (found: {repr(e2_value)})")
        else:
            formula_upper = e2_value.upper().replace(' ', '')
            has_sumifs = 'SUMIFS' in formula_upper
            has_b_range = 'B2:B365' in e2_value.upper().replace(' ', '')
            has_a_range = 'A2:A365' in e2_value.upper().replace(' ', '')

            if has_sumifs and has_b_range and has_a_range:
                print(f"PASS: Component 1 — E2 contains SUMIFS formula with correct data ranges: {e2_value} (0.4 pts)")
                total_score += 0.4
            else:
                print(f"FAIL: Component 1 — E2 formula incomplete. has_sumifs={has_sumifs}, has_b_range={has_b_range}, has_a_range={has_a_range}. Formula: {repr(e2_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: E2 formula uses correct Q4 2023 date criteria (0.3 pts)
    # The formula must include date bounds for Q4 2023: Oct 1 (month 10) and Dec 31 (month 12) of 2023.
    # Check for DATE(2023,10,1) and DATE(2023,12,31) or equivalent ">="/"<=" patterns.
    # This FAILS on initial (E2 is None) and PASSES on golden.
    try:
        e2_value = ws['E2'].value
        if e2_value is None or not isinstance(e2_value, str):
            print(f"FAIL: Component 2 — E2 is empty or not a formula")
        else:
            formula_no_space = e2_value.replace(' ', '')
            # Check for Q4 start: DATE(2023,10,1) or equivalent October 2023 reference
            has_oct_start = bool(re.search(r'DATE\s*\(\s*2023\s*,\s*10\s*,\s*1\s*\)', formula_no_space, re.IGNORECASE))
            # Check for Q4 end: DATE(2023,12,31) or equivalent December 2023 reference
            has_dec_end = bool(re.search(r'DATE\s*\(\s*2023\s*,\s*12\s*,\s*31\s*\)', formula_no_space, re.IGNORECASE))
            # Also check for ">=" and "<=" operators indicating date range
            has_gte = '>=' in formula_no_space or '">="' in formula_no_space
            has_lte = '<=' in formula_no_space or '"<="' in formula_no_space

            if has_oct_start and has_dec_end and has_gte and has_lte:
                print(f"PASS: Component 2 — Formula has correct Q4 2023 date bounds (Oct 1 to Dec 31, 2023) (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — Formula missing Q4 2023 date criteria. has_oct_start={has_oct_start}, has_dec_end={has_dec_end}, has_gte={has_gte}, has_lte={has_lte}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: E2 formula exists AND the underlying Q4 2023 data sums to ~1,234,780 (0.3 pts)
    # This compound check ensures:
    #   (a) E2 has a formula (not just empty — which is the initial state)
    #   (b) The data that formula would compute over sums to the correct expected value
    # This FAILS on initial (E2 is None → first condition fails immediately).
    try:
        e2_val = ws['E2'].value
        # Condition (a): E2 must have a formula — FAILS on initial file
        if e2_val is None or not isinstance(e2_val, str) or not e2_val.startswith('='):
            print(f"FAIL: Component 3 — E2 has no formula (found: {repr(e2_val)}), cannot verify result")
        else:
            # Condition (b): Manually sum Q4 2023 dates to confirm data integrity
            q4_start = datetime.datetime(2023, 10, 1)
            q4_end = datetime.datetime(2023, 12, 31, 23, 59, 59)
            q4_total = 0.0
            rows_counted = 0

            for row in range(2, 366):  # rows 2 to 365
                date_val = ws.cell(row=row, column=1).value
                amount_val = ws.cell(row=row, column=2).value
                if date_val is None or amount_val is None:
                    continue
                # Handle date as datetime object or date object
                if isinstance(date_val, datetime.datetime):
                    d = date_val
                elif isinstance(date_val, datetime.date):
                    d = datetime.datetime(date_val.year, date_val.month, date_val.day)
                else:
                    continue
                if q4_start <= d <= q4_end:
                    try:
                        q4_total += float(amount_val)
                        rows_counted += 1
                    except (ValueError, TypeError):
                        pass

            expected_total = 1234780
            tolerance = 1.0  # Allow up to 1.0 rounding difference
            pct_diff = abs(q4_total - expected_total) / expected_total if expected_total else 1.0

            if abs(q4_total - expected_total) <= tolerance or pct_diff < 0.01:
                print(f"PASS: Component 3 — Formula present AND Q4 2023 data total={q4_total:.2f} matches expected {expected_total} (rows: {rows_counted}) (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 3 — Q4 2023 data total={q4_total:.2f} does not match expected ~{expected_total} (rows: {rows_counted})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
