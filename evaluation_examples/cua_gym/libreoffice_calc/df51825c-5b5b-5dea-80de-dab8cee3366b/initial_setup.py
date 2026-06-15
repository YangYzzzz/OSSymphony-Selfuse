"""
Initial Setup: Create spreadsheet with strict whole-number validation on D10:D15
Task ID: calc_nrv_088
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_088'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Inventory ---
    ws = wb.active
    ws.title = 'Inventory'

    # Header styling
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    headers = ['Item Code', 'Product Name', 'Category', 'Quantity', 'Unit Price', 'Supplier']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 22

    # Realistic inventory data (rows 2-20)
    data = [
        ['INV-001', 'Wireless Bluetooth Headphones', 'Electronics', 245, 49.99, 'TechVision Supply Co.'],
        ['INV-002', 'Ergonomic Office Chair', 'Furniture', 38, 329.00, 'ComfortPlus Industries'],
        ['INV-003', 'Stainless Steel Water Bottle', 'Kitchen', 512, 24.50, 'EcoGreen Products'],
        ['INV-004', 'LED Desk Lamp (Adjustable)', 'Electronics', 167, 45.00, 'BrightHome Ltd.'],
        ['INV-005', 'Bamboo Cutting Board Set', 'Kitchen', 89, 32.75, 'NatureWorks Co.'],
        ['INV-006', 'USB-C Charging Hub (7-port)', 'Electronics', 324, 59.99, 'TechVision Supply Co.'],
        ['INV-007', 'Memory Foam Pillow', 'Bedding', 196, 39.95, 'DreamSoft Textiles'],
        ['INV-008', 'Cast Iron Skillet 12-inch', 'Kitchen', 73, 54.00, 'Heritage Cookware Inc.'],
        ['INV-009', 'Portable Bluetooth Speaker', 'Electronics', 410, 79.99, 'SoundWave Audio'],
        ['INV-010', 'Standing Desk Converter', 'Furniture', 52, 199.50, 'ComfortPlus Industries'],
        ['INV-011', 'Ceramic Coffee Mug Set (4)', 'Kitchen', 287, 28.00, 'ArtisanCraft Pottery'],
        ['INV-012', 'Noise-Cancelling Earbuds', 'Electronics', 183, 89.99, 'SoundWave Audio'],
        ['INV-013', 'Organic Cotton Throw Blanket', 'Bedding', 64, 55.00, 'DreamSoft Textiles'],
        ['INV-014', 'Adjustable Monitor Arm', 'Furniture', 128, 74.50, 'ErgoTech Solutions'],
        ['INV-015', 'Glass Food Storage Set (10pc)', 'Kitchen', 345, 42.99, 'EcoGreen Products'],
        ['INV-016', 'Mechanical Keyboard (Cherry MX)', 'Electronics', 91, 129.00, 'TechVision Supply Co.'],
        ['INV-017', 'Linen Duvet Cover Queen', 'Bedding', 47, 85.00, 'DreamSoft Textiles'],
        ['INV-018', 'Webcam HD 1080p', 'Electronics', 256, 64.99, 'TechVision Supply Co.'],
        ['INV-019', 'Bamboo Desk Organizer', 'Furniture', 178, 35.50, 'NatureWorks Co.'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 5:  # Unit Price column
                cell.number_format = '$#,##0.00'

    # --- Data Validation on D10:D15: strict whole number 1-1000 ---
    # This is the validation the user needs to remove and replace
    dv_strict = DataValidation(
        type='whole',
        operator='between',
        formula1='1',
        formula2='1000',
        allow_blank=False,
        showDropDown=False,
    )
    dv_strict.error = 'Please enter a whole number between 1 and 1000.'
    dv_strict.errorTitle = 'Invalid Quantity'
    dv_strict.prompt = 'Enter quantity (1-1000)'
    dv_strict.promptTitle = 'Quantity'
    dv_strict.showErrorMessage = True
    dv_strict.showInputMessage = True
    dv_strict.add('D10:D15')
    ws.add_data_validation(dv_strict)

    # --- Sheet 2: Suppliers ---
    ws2 = wb.create_sheet('Suppliers')
    supplier_headers = ['Supplier Name', 'Contact Email', 'Phone', 'Region']
    for col, h in enumerate(supplier_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    suppliers = [
        ['TechVision Supply Co.', 'orders@techvision.com', '(415) 555-0142', 'West Coast'],
        ['ComfortPlus Industries', 'sales@comfortplus.net', '(312) 555-0198', 'Midwest'],
        ['EcoGreen Products', 'info@ecogreen.com', '(503) 555-0267', 'Pacific NW'],
        ['BrightHome Ltd.', 'contact@brighthome.co', '(617) 555-0321', 'Northeast'],
        ['NatureWorks Co.', 'hello@natureworks.org', '(206) 555-0189', 'Pacific NW'],
        ['DreamSoft Textiles', 'orders@dreamsoft.com', '(919) 555-0234', 'Southeast'],
        ['Heritage Cookware Inc.', 'sales@heritagecook.com', '(615) 555-0156', 'Southeast'],
        ['SoundWave Audio', 'biz@soundwave.io', '(213) 555-0278', 'West Coast'],
        ['ArtisanCraft Pottery', 'studio@artisancraft.com', '(802) 555-0145', 'Northeast'],
        ['ErgoTech Solutions', 'support@ergotech.com', '(408) 555-0312', 'West Coast'],
    ]
    for r, row_data in enumerate(suppliers, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    ws2.column_dimensions['A'].width = 26
    ws2.column_dimensions['B'].width = 28
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 14

    # Freeze header row on both sheets
    ws.freeze_panes = 'A2'
    ws2.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
