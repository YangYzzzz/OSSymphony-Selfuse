"""
Reward Script: Modify 'Heading 1' cell style to Arial 14pt bold with light gray background
Task ID: calc_gfl_041
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): Font size is 14pt on all 5 section headers (initial has 18pt)
  Component 2 (0.40): Background fill is light gray on all 5 section headers (initial has light blue)
  Component 3 (0.25): All properties correct together: Arial, 14pt, bold, light gray on all headers
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_041'

# Section header rows per task context
HEADER_ROWS = [1, 8, 15, 22, 29]
EXPECTED_HEADERS = ['Personnel Costs', 'Technology', 'Marketing', 'Operations', 'Contingency']


def is_light_gray(cell):
    """Check if cell background is a light gray color (R~=G~=B, in 190-230 range)."""
    try:
        if cell.fill.patternType != 'solid':
            return False
        rgb_str = str(cell.fill.fgColor.rgb)
        if len(rgb_str) == 8:
            r, g, b = int(rgb_str[2:4], 16), int(rgb_str[4:6], 16), int(rgb_str[6:8], 16)
        elif len(rgb_str) == 6:
            r, g, b = int(rgb_str[0:2], 16), int(rgb_str[2:4], 16), int(rgb_str[4:6], 16)
        else:
            return False
        return (abs(r - g) <= 10 and abs(g - b) <= 10 and abs(r - b) <= 10 and
                190 <= r <= 230 and 190 <= g <= 230 and 190 <= b <= 230)
    except Exception:
        return False


def is_size_14(cell):
    """Check if font size is 14pt (with small tolerance)."""
    try:
        return cell.font.size is not None and abs(float(cell.font.size) - 14.0) < 0.5
    except Exception:
        return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Budget sheet exists
    if 'Budget' not in wb.sheetnames:
        print("FAIL: 'Budget' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Budget']

    # Precondition: Section headers still have correct text values
    for i, row in enumerate(HEADER_ROWS):
        val = ws.cell(row=row, column=1).value
        if val != EXPECTED_HEADERS[i]:
            print(f"FAIL: Row {row} expected '{EXPECTED_HEADERS[i]}', found '{val}'")
            print("REWARD: 0.0")
            return 0.0

    # Component 1: Font size is 14pt on all 5 section headers (0.35 points)
    # Initial state has 18pt; golden has 14pt. This is a task-introduced change.
    try:
        count = 0
        for row in HEADER_ROWS:
            cell = ws.cell(row=row, column=1)
            if is_size_14(cell):
                count += 1
            else:
                print(f"  Row {row}: font size = {cell.font.size!r} (expected 14)")

        if count == 5:
            print(f"PASS: Component 1 — All 5 headers have 14pt font (0.35 pts)")
            total_score += 0.35
        elif count > 0:
            partial = round(0.35 * count / 5, 3)
            print(f"PARTIAL: Component 1 — {count}/5 headers have 14pt ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No headers have 14pt font")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Background fill is light gray on all 5 section headers (0.40 points)
    # Initial state has FFDCE6F1 (light blue); golden has FFD9D9D9 (light gray).
    try:
        count = 0
        for row in HEADER_ROWS:
            cell = ws.cell(row=row, column=1)
            if is_light_gray(cell):
                count += 1
            else:
                try:
                    rgb = cell.fill.fgColor.rgb
                except Exception:
                    rgb = 'N/A'
                print(f"  Row {row}: fill color = {rgb} (expected light gray)")

        if count == 5:
            print(f"PASS: Component 2 — All 5 headers have light gray background (0.40 pts)")
            total_score += 0.40
        elif count > 0:
            partial = round(0.40 * count / 5, 3)
            print(f"PARTIAL: Component 2 — {count}/5 headers have light gray ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No headers have light gray background")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: All properties correct together on all 5 headers (0.25 points)
    # Compound check: Arial font, 14pt, bold, light gray background — all at once.
    # This verifies the complete style modification, not just individual parts.
    try:
        count = 0
        for row in HEADER_ROWS:
            cell = ws.cell(row=row, column=1)
            name_ok = cell.font.name and cell.font.name.lower() == 'arial'
            size_ok = is_size_14(cell)
            bold_ok = cell.font.bold is True
            gray_ok = is_light_gray(cell)

            if name_ok and size_ok and bold_ok and gray_ok:
                count += 1
            else:
                print(f"  Row {row}: arial={name_ok}, 14pt={size_ok}, bold={bold_ok}, gray={gray_ok}")

        if count == 5:
            print(f"PASS: Component 3 — All 5 headers have complete correct style (0.25 pts)")
            total_score += 0.25
        elif count > 0:
            partial = round(0.25 * count / 5, 3)
            print(f"PARTIAL: Component 3 — {count}/5 headers fully correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No headers have complete correct style")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved LibreOffice state before verification
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(1.0)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
