"""
Initial Setup: Modify the built-in 'Heading 1' cell style to use Arial 14pt bold
with a light gray background and apply it to all section headers.
Task ID: calc_gfl_041
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_041'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Budget'

    # Define the default 'Heading 1' look: blue large font (mimics LibreOffice default)
    heading1_font = Font(name='Liberation Sans', size=18, bold=True, color='003366')
    heading1_fill = PatternFill(start_color='FFDCE6F1', end_color='FFDCE6F1', fill_type='solid')  # light blue tint

    # Section headers at rows 1, 8, 15, 22, 29
    section_headers = {
        1: 'Personnel Costs',
        8: 'Technology',
        15: 'Marketing',
        22: 'Operations',
        29: 'Contingency',
    }

    # Column headers for data rows (placed in the row after each section header)
    data_columns = ['Category', 'Q1 Budget', 'Q2 Budget', 'Q3 Budget', 'Q4 Budget', 'Annual Total']
    col_header_font = Font(name='Calibri', size=11, bold=True)

    # --- Personnel Costs (rows 1-7) ---
    ws.cell(row=1, column=1, value='Personnel Costs')
    for c, h in enumerate(data_columns, 1):
        ws.cell(row=2, column=c, value=h).font = col_header_font
    personnel_data = [
        ['Salaries - Engineering', 125000, 125000, 130000, 130000, 510000],
        ['Salaries - Marketing', 85000, 85000, 87000, 87000, 344000],
        ['Benefits & Insurance', 42000, 42000, 43500, 43500, 171000],
        ['Training & Development', 8000, 12000, 8000, 15000, 43000],
        ['Recruitment Costs', 15000, 10000, 20000, 10000, 55000],
    ]
    for r, row_data in enumerate(personnel_data, 3):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # --- Technology (rows 8-14) ---
    ws.cell(row=8, column=1, value='Technology')
    for c, h in enumerate(data_columns, 1):
        ws.cell(row=9, column=c, value=h).font = col_header_font
    tech_data = [
        ['Cloud Infrastructure (AWS)', 18000, 19500, 21000, 22000, 80500],
        ['Software Licenses', 12000, 12000, 14000, 14000, 52000],
        ['Hardware Refresh', 25000, 5000, 5000, 30000, 65000],
        ['Cybersecurity Tools', 8500, 8500, 9000, 9000, 35000],
        ['IT Support Contracts', 6000, 6000, 6000, 6000, 24000],
    ]
    for r, row_data in enumerate(tech_data, 10):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # --- Marketing (rows 15-21) ---
    ws.cell(row=15, column=1, value='Marketing')
    for c, h in enumerate(data_columns, 1):
        ws.cell(row=16, column=c, value=h).font = col_header_font
    marketing_data = [
        ['Digital Advertising', 22000, 28000, 35000, 40000, 125000],
        ['Content Production', 8000, 8000, 10000, 10000, 36000],
        ['Trade Shows & Events', 15000, 5000, 25000, 10000, 55000],
        ['PR & Communications', 5000, 5000, 6000, 6000, 22000],
        ['Market Research', 10000, 7000, 10000, 8000, 35000],
    ]
    for r, row_data in enumerate(marketing_data, 17):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # --- Operations (rows 22-28) ---
    ws.cell(row=22, column=1, value='Operations')
    for c, h in enumerate(data_columns, 1):
        ws.cell(row=23, column=c, value=h).font = col_header_font
    ops_data = [
        ['Office Lease', 30000, 30000, 30000, 30000, 120000],
        ['Utilities & Maintenance', 4500, 4000, 5000, 5500, 19000],
        ['Office Supplies', 2000, 2000, 2500, 3000, 9500],
        ['Travel & Transportation', 8000, 12000, 10000, 15000, 45000],
        ['Insurance Premiums', 7500, 7500, 7500, 7500, 30000],
    ]
    for r, row_data in enumerate(ops_data, 24):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # --- Contingency (rows 29-35) ---
    ws.cell(row=29, column=1, value='Contingency')
    for c, h in enumerate(data_columns, 1):
        ws.cell(row=30, column=c, value=h).font = col_header_font
    contingency_data = [
        ['Emergency Reserve', 15000, 15000, 15000, 15000, 60000],
        ['Project Overruns', 10000, 10000, 12000, 12000, 44000],
        ['Regulatory Compliance', 5000, 5000, 7000, 7000, 24000],
        ['Currency Fluctuation Buffer', 3000, 3000, 4000, 4000, 14000],
        ['Miscellaneous', 5000, 5000, 5000, 5000, 20000],
    ]
    for r, row_data in enumerate(contingency_data, 31):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Apply default 'Heading 1' appearance to section headers (blue, large)
    for row_num in section_headers:
        cell = ws.cell(row=row_num, column=1)
        cell.font = heading1_font
        cell.fill = heading1_fill
        cell.alignment = Alignment(vertical='center')

    # Set column widths for readability
    ws.column_dimensions['A'].width = 30
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col_letter].width = 15

    # Format currency columns
    for r in range(1, 36):
        for c in range(2, 7):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)) and cell.value:
                cell.number_format = '$#,##0'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
