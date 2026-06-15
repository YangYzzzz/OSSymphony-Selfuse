"""
Initial Setup: Create Delivery_Schedule.xlsx with 34 orders, Column F formatted as date but empty, no validation.
Task ID: calc_gcv_089
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'Delivery_Schedule'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Deliveries"

    # Headers
    headers = ["Order ID", "Product", "Quantity", "Origin", "Destination", "Delivery Date"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 34 orders with realistic data
    orders = [
        ["ORD-1001", "Samsung Galaxy S24", 15, "Seoul", "Los Angeles"],
        ["ORD-1002", "MacBook Pro 16-inch", 8, "Shenzhen", "New York"],
        ["ORD-1003", "Sony WH-1000XM5 Headphones", 50, "Tokyo", "London"],
        ["ORD-1004", "Dell XPS 15 Laptop", 12, "Austin", "Toronto"],
        ["ORD-1005", "Canon EOS R6 Camera", 6, "Osaka", "Berlin"],
        ["ORD-1006", "LG OLED 65\" TV", 4, "Busan", "Sydney"],
        ["ORD-1007", "Dyson V15 Vacuum", 20, "Singapore", "Dubai"],
        ["ORD-1008", "Apple iPad Air", 30, "Shenzhen", "Paris"],
        ["ORD-1009", "Bose QuietComfort Earbuds", 45, "Boston", "Mumbai"],
        ["ORD-1010", "Nintendo Switch OLED", 25, "Kyoto", "Madrid"],
        ["ORD-1011", "HP LaserJet Printer", 10, "Palo Alto", "São Paulo"],
        ["ORD-1012", "Logitech MX Master Mouse", 60, "Lausanne", "Chicago"],
        ["ORD-1013", "Google Pixel 8 Pro", 18, "Mountain View", "Amsterdam"],
        ["ORD-1014", "Philips Hue Starter Kit", 35, "Eindhoven", "Stockholm"],
        ["ORD-1015", "Kindle Paperwhite", 40, "Seattle", "Rome"],
        ["ORD-1016", "JBL Charge 5 Speaker", 22, "Los Angeles", "Bangkok"],
        ["ORD-1017", "Nikon Z8 Camera", 5, "Tokyo", "Vancouver"],
        ["ORD-1018", "Razer Blade 16 Laptop", 7, "San Francisco", "Shanghai"],
        ["ORD-1019", "Garmin Fenix 7 Watch", 14, "Olathe", "Zurich"],
        ["ORD-1020", "Sonos Arc Soundbar", 9, "Santa Barbara", "Melbourne"],
        ["ORD-1021", "Lenovo ThinkPad X1 Carbon", 16, "Beijing", "Warsaw"],
        ["ORD-1022", "GoPro Hero 12 Black", 28, "San Mateo", "Cape Town"],
        ["ORD-1023", "Breville Espresso Machine", 11, "Sydney", "Milan"],
        ["ORD-1024", "Anker PowerCore 26800", 55, "Shenzhen", "Mexico City"],
        ["ORD-1025", "Microsoft Surface Pro 10", 13, "Redmond", "Helsinki"],
        ["ORD-1026", "Roomba j9+ Robot Vacuum", 17, "Bedford", "Vienna"],
        ["ORD-1027", "Sennheiser Momentum 4", 32, "Hanover", "Lisbon"],
        ["ORD-1028", "Fujifilm X-T5 Camera", 8, "Tokyo", "Dublin"],
        ["ORD-1029", "Instant Pot Duo Plus", 42, "Ottawa", "Jakarta"],
        ["ORD-1030", "Asus ROG Ally Handheld", 19, "Taipei", "Prague"],
        ["ORD-1031", "Weber Genesis Grill", 3, "Palatine", "Copenhagen"],
        ["ORD-1032", "DJI Air 3 Drone", 10, "Shenzhen", "Athens"],
        ["ORD-1033", "Fitbit Sense 2", 26, "San Francisco", "Oslo"],
        ["ORD-1034", "Yamaha YAS-209 Soundbar", 15, "Hamamatsu", "Budapest"],
    ]

    data_font = Font(name="Calibri", size=11)
    data_align = Alignment(vertical="center")

    for r, row_data in enumerate(orders, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

        # Column F (Delivery Date) - formatted as date but left EMPTY
        f_cell = ws.cell(row=r, column=6)
        f_cell.number_format = 'YYYY-MM-DD'
        f_cell.font = data_font
        f_cell.alignment = data_align
        f_cell.border = thin_border

    # Set column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 16

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
