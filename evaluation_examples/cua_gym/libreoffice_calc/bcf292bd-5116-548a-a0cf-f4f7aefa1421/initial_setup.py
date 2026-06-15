"""
Initial Setup: Create a spreadsheet with deal data for top 3 extraction task
Task ID: calc_sales_073
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_073'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'TopDeals'

    # --- Headers ---
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal='center')
    white_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")

    # Source data headers
    for col, header in enumerate(['Deal', 'Value'], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Result area headers
    for col, header in [(4, 'Rank'), (5, 'Deal Name'), (6, 'Value')]:
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Deal data (A2:B8) ---
    deals = [
        ('Alpha', 75000),
        ('Beta', 210000),
        ('Gamma', 45000),
        ('Delta', 320000),
        ('Epsilon', 150000),
        ('Zeta', 280000),
        ('Eta', 95000),
    ]

    for r, (name, value) in enumerate(deals, 2):
        ws.cell(row=r, column=1, value=name)
        val_cell = ws.cell(row=r, column=2, value=value)
        val_cell.number_format = '#,##0'

    # --- Rank numbers (D2:D4) ---
    for r, rank in enumerate([1, 2, 3], 2):
        cell = ws.cell(row=r, column=4, value=rank)
        cell.alignment = Alignment(horizontal='center')

    # E2:E4 and F2:F4 left EMPTY - task is to fill these with formulas

    # --- Column widths ---
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 4   # spacer
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
