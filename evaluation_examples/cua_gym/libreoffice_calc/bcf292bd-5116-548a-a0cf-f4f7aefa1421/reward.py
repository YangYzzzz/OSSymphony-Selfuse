"""
Reward Script: Create array formula (CSE) to find top 3 deals by value
Task ID: calc_sales_073
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): F2:F4 contain LARGE formulas for top 3 values
  Component 2 (0.35): E2:E4 contain INDEX/MATCH formulas for deal names
  Component 3 (0.15): LARGE formulas reference correct range and ranks 1,2,3
  Component 4 (0.15): INDEX/MATCH formulas reference correct range and use LARGE
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_073'


def normalize_formula(f):
    """Normalize a formula string for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: TopDeals sheet must exist
    if 'TopDeals' not in wb.sheetnames:
        print("CRITICAL: 'TopDeals' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['TopDeals']

    # Component 1: F2:F4 contain LARGE formulas for top 3 values (0.35 points)
    # The golden state should have LARGE(B2:B8,N) formulas in F2, F3, F4
    try:
        large_count = 0
        for cell_ref in ['F2', 'F3', 'F4']:
            val = ws[cell_ref].value
            if val is not None and isinstance(val, str):
                nf = normalize_formula(val)
                # Check that the formula contains LARGE function
                if '=LARGE(' in nf or 'LARGE(' in nf:
                    large_count += 1
                    print(f"  PASS: {cell_ref} contains LARGE formula: {val}")
                else:
                    print(f"  FAIL: {cell_ref} has formula but no LARGE: {val}")
            else:
                print(f"  FAIL: {cell_ref} is empty or not a formula (value: {repr(val)})")

        if large_count == 3:
            print(f"PASS: Component 1 — All 3 cells F2:F4 have LARGE formulas (0.35 pts)")
            total_score += 0.35
        elif large_count > 0:
            partial = round(0.35 * large_count / 3, 2)
            print(f"PARTIAL: Component 1 — {large_count}/3 cells have LARGE formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No LARGE formulas found in F2:F4")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: E2:E4 contain INDEX/MATCH formulas for deal names (0.35 points)
    # The golden state should have INDEX(...MATCH(LARGE(...))) in E2, E3, E4
    try:
        index_match_count = 0
        for cell_ref in ['E2', 'E3', 'E4']:
            val = ws[cell_ref].value
            if val is not None and isinstance(val, str):
                nf = normalize_formula(val)
                # Check for INDEX and MATCH combination
                if 'INDEX(' in nf and 'MATCH(' in nf:
                    index_match_count += 1
                    print(f"  PASS: {cell_ref} contains INDEX/MATCH formula: {val}")
                else:
                    print(f"  FAIL: {cell_ref} has formula but no INDEX/MATCH: {val}")
            else:
                print(f"  FAIL: {cell_ref} is empty or not a formula (value: {repr(val)})")

        if index_match_count == 3:
            print(f"PASS: Component 2 — All 3 cells E2:E4 have INDEX/MATCH formulas (0.35 pts)")
            total_score += 0.35
        elif index_match_count > 0:
            partial = round(0.35 * index_match_count / 3, 2)
            print(f"PARTIAL: Component 2 — {index_match_count}/3 cells have INDEX/MATCH formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No INDEX/MATCH formulas found in E2:E4")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: LARGE formulas reference correct range (B2:B8) and ranks 1,2,3 (0.15 points)
    try:
        correct_large = 0
        expected_ranks = {'F2': '1', 'F3': '2', 'F4': '3'}
        for cell_ref, expected_rank in expected_ranks.items():
            val = ws[cell_ref].value
            if val is not None and isinstance(val, str):
                nf = normalize_formula(val)
                # Check for LARGE(B2:B8,N) pattern
                pattern = f'LARGE(B2:B8,{expected_rank})'
                if pattern in nf:
                    correct_large += 1
                    print(f"  PASS: {cell_ref} has correct LARGE formula with rank {expected_rank}")
                else:
                    # Also accept LARGE($B$2:$B$8,N) with absolute refs
                    nf_no_dollar = nf.replace('$', '')
                    if pattern in nf_no_dollar:
                        correct_large += 1
                        print(f"  PASS: {cell_ref} has correct LARGE formula (absolute refs) with rank {expected_rank}")
                    else:
                        print(f"  FAIL: {cell_ref} LARGE formula doesn't match expected pattern: {val}")
            else:
                print(f"  FAIL: {cell_ref} is empty or not a formula")

        if correct_large == 3:
            print(f"PASS: Component 3 — All LARGE formulas have correct range and ranks (0.15 pts)")
            total_score += 0.15
        elif correct_large > 0:
            partial = round(0.15 * correct_large / 3, 2)
            print(f"PARTIAL: Component 3 — {correct_large}/3 correct LARGE formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No correctly structured LARGE formulas")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: INDEX/MATCH formulas use LARGE and correct references (0.15 points)
    # E2:E4 should reference A2:A8 (names), B2:B8 (values), and use LARGE(B2:B8,N)
    try:
        correct_im = 0
        expected_ranks_e = {'E2': '1', 'E3': '2', 'E4': '3'}
        for cell_ref, expected_rank in expected_ranks_e.items():
            val = ws[cell_ref].value
            if val is not None and isinstance(val, str):
                nf = normalize_formula(val)
                nf_no_dollar = nf.replace('$', '')
                # Check that formula uses INDEX with A range and MATCH with LARGE
                has_index_a = 'INDEX(A2:A8' in nf_no_dollar or 'INDEX(A:A' in nf_no_dollar
                has_large_rank = f'LARGE(B2:B8,{expected_rank})' in nf_no_dollar
                has_match = 'MATCH(' in nf_no_dollar

                if has_index_a and has_large_rank and has_match:
                    correct_im += 1
                    print(f"  PASS: {cell_ref} INDEX/MATCH uses correct ranges and LARGE rank {expected_rank}")
                else:
                    # Be more lenient - just check INDEX, MATCH, and LARGE are present with correct rank
                    if 'INDEX(' in nf and 'MATCH(' in nf and f'LARGE(' in nf and f',{expected_rank})' in nf_no_dollar:
                        correct_im += 1
                        print(f"  PASS: {cell_ref} INDEX/MATCH/LARGE structure correct with rank {expected_rank}")
                    else:
                        print(f"  FAIL: {cell_ref} formula structure doesn't match expected: {val}")
            else:
                print(f"  FAIL: {cell_ref} is empty or not a formula")

        if correct_im == 3:
            print(f"PASS: Component 4 — All INDEX/MATCH formulas correctly structured (0.15 pts)")
            total_score += 0.15
        elif correct_im > 0:
            partial = round(0.15 * correct_im / 3, 2)
            print(f"PARTIAL: Component 4 — {correct_im}/3 correct INDEX/MATCH formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — No correctly structured INDEX/MATCH formulas")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
