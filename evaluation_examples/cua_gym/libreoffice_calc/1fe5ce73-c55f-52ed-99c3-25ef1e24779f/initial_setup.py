"""
Initial Setup: Format time values to show elapsed hours and minutes beyond 24 hours
Task ID: calc_lf_081
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_081'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Timesheet'

    # Headers
    ws['A1'] = 'Task'
    ws['B1'] = 'Duration'

    # Data rows with raw time values (fraction of a day)
    # 36 hours 45 minutes = 36.75 / 24 = 1.53125
    ws['A2'] = 'Project A'
    ws['B2'] = 1.53125

    # 8 hours 30 minutes = 8.5 / 24 = 0.354166666666667
    ws['A3'] = 'Project B'
    ws['B3'] = 8.5 / 24

    # 45 hours 15 minutes = 45.25 / 24 = 1.885416666666667
    ws['A4'] = 'Total'
    ws['B4'] = 45.25 / 24

    # Leave B2:B4 in General format (no [HH]:MM applied)
    # They will display as decimal numbers like 1.53125

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Open in LibreOffice Calc for GUI-ready state
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
