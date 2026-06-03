"""
Reward Script: Format time values with [HH]:MM format for timesheet total
Task ID: calc_lf_081
Domain: libreoffice_calc
Scoring:
  - Component 1: B2 number format is [HH]:MM (0.35 pts)
  - Component 2: B3 number format is [HH]:MM (0.30 pts)
  - Component 3: B4 number format is [HH]:MM (0.35 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_081'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    The task requires applying '[HH]:MM' number format to cells B2:B4
    on the 'Timesheet' sheet. This format displays elapsed hours beyond 24.
    """
    total_score = 0.0

    # Precondition: load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Timesheet' sheet exists
    if 'Timesheet' not in wb.sheetnames:
        print(f"CRITICAL: 'Timesheet' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Timesheet']

    # Precondition: B2:B4 have non-None time values (data integrity gate)
    for coord in ['B2', 'B3', 'B4']:
        val = ws[coord].value
        if val is None:
            print(f"CRITICAL: {coord} has no value — data corrupted")
            print("REWARD: 0.0")
            return 0.0

    # The acceptable format patterns for [HH]:MM elapsed time
    # openpyxl may store it as '[HH]:MM', '[h]:mm', '[hh]:mm', etc.
    # We normalize to lowercase for comparison
    def is_elapsed_time_format(fmt):
        """Check if number format is an elapsed hours:minutes format."""
        if fmt is None:
            return False
        fmt_lower = fmt.lower().strip()
        # Must have square brackets around h/hh and :mm
        # Valid patterns: [h]:mm, [hh]:mm, [h]:mm:ss, [hh]:mm:ss
        if '[h]' in fmt_lower or '[hh]' in fmt_lower:
            if ':mm' in fmt_lower:
                return True
        return False

    # Component 1: B2 has [HH]:MM number format (0.35 points)
    # This checks the format change — initial has 'General', golden has '[HH]:MM'
    try:
        fmt_b2 = ws['B2'].number_format
        if is_elapsed_time_format(fmt_b2):
            print(f"PASS: Component 1 — B2 has elapsed time format: '{fmt_b2}' (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 1 — B2 expected [HH]:MM format, found: '{fmt_b2}'")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: B3 has [HH]:MM number format (0.30 points)
    try:
        fmt_b3 = ws['B3'].number_format
        if is_elapsed_time_format(fmt_b3):
            print(f"PASS: Component 2 — B3 has elapsed time format: '{fmt_b3}' (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 2 — B3 expected [HH]:MM format, found: '{fmt_b3}'")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: B4 has [HH]:MM number format (0.35 points)
    try:
        fmt_b4 = ws['B4'].number_format
        if is_elapsed_time_format(fmt_b4):
            print(f"PASS: Component 3 — B4 has elapsed time format: '{fmt_b4}' (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 3 — B4 expected [HH]:MM format, found: '{fmt_b4}'")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint: persist app state then verify
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
