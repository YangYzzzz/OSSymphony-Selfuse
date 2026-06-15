"""
Initial Setup: Create MonthlySales spreadsheet with 360 transaction rows
Task ID: calc_pivot_072
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_072'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

# Monthly sales targets that produce the required cumulative percentages
MONTHLY_TARGETS = {
    1: 9960, 2: 10560, 3: 10440, 4: 10080, 5: 10440, 6: 10320,
    7: 10320, 8: 10440, 9: 10080, 10: 10200, 11: 6960, 12: 10200
}

PRODUCTS = [
    'Wireless Keyboard', 'USB-C Hub', 'Monitor Stand', 'Desk Lamp',
    'Webcam HD', 'Noise-Cancel Headphones', 'Ergonomic Mouse',
    'Laptop Sleeve', 'Cable Organizer', 'Screen Protector',
    'Portable SSD', 'Bluetooth Speaker', 'Phone Charger',
    'Tablet Case', 'HDMI Cable', 'Power Strip', 'Mousepad XL',
    'Webcam Light', 'USB Microphone', 'Docking Station'
]

MONTH_NAMES = [
    'January', 'February', 'March', 'April', 'May', 'June',
    'July', 'August', 'September', 'October', 'November', 'December'
]


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'MonthlySales'

    # --- Headers ---
    headers = ['TxnID', 'Month', 'Product', 'Sales']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Generate 360 rows: 30 per month ---
    # Distribute monthly target across 30 transactions
    txn_id = 1
    row_num = 2
    for month in range(1, 13):
        target = MONTHLY_TARGETS[month]
        # Generate 30 random sales amounts summing to target
        # Use a simple approach: generate 29 random amounts, last one is remainder
        amounts = []
        remaining = target
        for i in range(29):
            # Average per remaining transaction
            avg = remaining / (30 - i)
            amt = round(random.uniform(avg * 0.4, avg * 1.6), 2)
            amt = max(50, min(amt, remaining - 50 * (29 - i)))  # ensure enough left
            amt = round(amt, 2)
            amounts.append(amt)
            remaining -= amt
        amounts.append(round(remaining, 2))
        random.shuffle(amounts)

        for amt in amounts:
            # Date: random day in the month
            max_day = 28 if month == 2 else (30 if month in [4, 6, 9, 11] else 31)
            day = random.randint(1, max_day)
            date_val = datetime(2024, month, day)
            product = random.choice(PRODUCTS)

            ws.cell(row=row_num, column=1, value=txn_id)
            ws.cell(row=row_num, column=2, value=date_val)
            ws.cell(row=row_num, column=2).number_format = 'yyyy-mm-dd'
            ws.cell(row=row_num, column=3, value=product)
            ws.cell(row=row_num, column=4, value=round(amt, 2))
            ws.cell(row=row_num, column=4).number_format = '#,##0.00'

            txn_id += 1
            row_num += 1

    # --- Column widths ---
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
