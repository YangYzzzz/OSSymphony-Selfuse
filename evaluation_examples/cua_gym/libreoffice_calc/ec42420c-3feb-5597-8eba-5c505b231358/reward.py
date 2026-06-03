"""
Reward Script: Running percentage total pivot table of sales by month
Task ID: calc_pivot_072
Domain: libreoffice_calc
Scoring:
  Component 1: PivotTable sheet exists (0.15 pts)
  Component 2: Month labels correct in A2:A13 (0.15 pts)
  Component 3: Sum of Sales values correct per month in B2:B13 (0.25 pts)
  Component 4: Running % Total values correct and cumulative in C2:C13 (0.30 pts)
  Component 5: Grand Total row present with correct totals (0.15 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_072'

# Expected month labels in order
EXPECTED_MONTHS = [
    'January', 'February', 'March', 'April', 'May', 'June',
    'July', 'August', 'September', 'October', 'November', 'December'
]

# Expected monthly sales sums from golden env
EXPECTED_MONTHLY_SALES = {
    'January': 9960,
    'February': 10560,
    'March': 10440,
    'April': 10080,
    'May': 10440,
    'June': 10320,
    'July': 10320,
    'August': 10440,
    'September': 10080,
    'October': 10200,
    'November': 6960,
    'December': 10200,
}

# Expected running % total values (cumulative percentage)
EXPECTED_RUNNING_PCT = {
    'January': 0.083,
    'February': 0.171,
    'March': 0.258,
    'April': 0.342,
    'May': 0.429,
    'June': 0.515,
    'July': 0.601,
    'August': 0.688,
    'September': 0.772,
    'October': 0.857,
    'November': 0.915,
    'December': 1.0,
}

EXPECTED_GRAND_TOTAL = 120000


def persist_app_state(domain):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: PivotTable sheet exists (0.15 points)
    # This FAILS on initial (only MonthlySales) and PASSES on golden
    try:
        pivot_ws = None
        for name in wb.sheetnames:
            if 'pivot' in name.lower():
                pivot_ws = wb[name]
                break
        if pivot_ws is not None:
            print(f"PASS: Component 1 — PivotTable sheet found: '{pivot_ws.title}' (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — No sheet with 'pivot' in name found. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0  # No pivot sheet means nothing else can be verified
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 2: Month labels correct in rows (0.15 points)
    # Check that 12 month names appear in column A in chronological order
    try:
        # Find the month column - look in first few columns for month names
        month_col = None
        header_row = None
        for col_idx in range(1, pivot_ws.max_column + 1):
            for row_idx in range(1, min(3, pivot_ws.max_row + 1)):
                val = pivot_ws.cell(row=row_idx, column=col_idx).value
                if val and isinstance(val, str) and 'month' in val.lower():
                    month_col = col_idx
                    header_row = row_idx
                    break
            if month_col:
                break

        if not month_col:
            # Fallback: assume column A with month names starting at row 2
            month_col = 1
            header_row = 1

        # Read month values starting from row after header
        data_start = header_row + 1
        found_months = []
        for r in range(data_start, pivot_ws.max_row + 1):
            val = pivot_ws.cell(row=r, column=month_col).value
            if val and isinstance(val, str) and val.strip().lower() not in ('grand total', 'total', ''):
                found_months.append(val.strip())

        # Check month names match expected (case-insensitive)
        correct_months = 0
        for i, expected in enumerate(EXPECTED_MONTHS):
            if i < len(found_months) and found_months[i].lower() == expected.lower():
                correct_months += 1

        if correct_months == 12:
            print(f"PASS: Component 2 — All 12 month labels correct in order (0.15 pts)")
            total_score += 0.15
        elif correct_months >= 6:
            partial = 0.15 * (correct_months / 12)
            print(f"PARTIAL: Component 2 — {correct_months}/12 month labels correct ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {correct_months}/12 month labels correct. Found: {found_months[:5]}...")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Sum of Sales values correct per month (0.25 points)
    # Check that column B (or sales column) has correct monthly sums
    try:
        # Find sales column - look for header containing 'sales' or 'sum'
        sales_col = None
        for col_idx in range(1, pivot_ws.max_column + 1):
            val = pivot_ws.cell(row=header_row, column=col_idx).value
            if val and isinstance(val, str) and ('sales' in val.lower() or 'sum' in val.lower()):
                sales_col = col_idx
                break

        if not sales_col:
            # Fallback: column B
            sales_col = 2

        correct_sales = 0
        for i, month_name in enumerate(EXPECTED_MONTHS):
            row_idx = data_start + i
            cell_val = pivot_ws.cell(row=row_idx, column=sales_col).value
            expected_val = EXPECTED_MONTHLY_SALES[month_name]
            if cell_val is not None:
                try:
                    actual = float(cell_val)
                    # Allow tolerance of 1.0 for rounding
                    if abs(actual - expected_val) <= 1.0:
                        correct_sales += 1
                    else:
                        print(f"  INFO: {month_name} sales: expected ~{expected_val}, got {actual}")
                except (ValueError, TypeError):
                    print(f"  INFO: {month_name} sales: non-numeric value {cell_val!r}")

        if correct_sales == 12:
            print(f"PASS: Component 3 — All 12 monthly sales sums correct (0.25 pts)")
            total_score += 0.25
        elif correct_sales >= 6:
            partial = 0.25 * (correct_sales / 12)
            print(f"PARTIAL: Component 3 — {correct_sales}/12 monthly sales correct ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {correct_sales}/12 monthly sales correct")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Running % Total values correct and cumulative (0.30 points)
    # This is the core task requirement - cumulative percentage
    try:
        # Find running % column - look for header containing '%' or 'running' or 'cumul'
        pct_col = None
        for col_idx in range(1, pivot_ws.max_column + 1):
            val = pivot_ws.cell(row=header_row, column=col_idx).value
            if val and isinstance(val, str) and ('%' in val.lower() or 'running' in val.lower() or 'cumul' in val.lower()):
                pct_col = col_idx
                break

        if not pct_col:
            # Fallback: column C
            pct_col = 3

        correct_pct = 0
        monotonic_violations = 0
        prev_val = 0.0

        for i, month_name in enumerate(EXPECTED_MONTHS):
            row_idx = data_start + i
            cell_val = pivot_ws.cell(row=row_idx, column=pct_col).value
            expected_pct = EXPECTED_RUNNING_PCT[month_name]

            if cell_val is not None:
                try:
                    actual = float(cell_val)
                    # Allow tolerance of 0.005 (0.5%) for rounding
                    if abs(actual - expected_pct) <= 0.005:
                        correct_pct += 1
                    else:
                        print(f"  INFO: {month_name} running %: expected ~{expected_pct}, got {actual}")

                    # Check monotonically increasing
                    if actual < prev_val - 0.001:
                        monotonic_violations += 1
                    prev_val = actual
                except (ValueError, TypeError):
                    print(f"  INFO: {month_name} running %: non-numeric value {cell_val!r}")
                    monotonic_violations += 1

        # Sub-component 4a: Values correct (0.20 pts)
        if correct_pct == 12:
            print(f"PASS: Component 4a — All 12 running % values correct (0.20 pts)")
            total_score += 0.20
        elif correct_pct >= 6:
            partial = 0.20 * (correct_pct / 12)
            print(f"PARTIAL: Component 4a — {correct_pct}/12 running % values correct ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4a — Only {correct_pct}/12 running % values correct")

        # Sub-component 4b: Monotonically increasing and Dec == 100% (0.10 pts)
        dec_val = pivot_ws.cell(row=data_start + 11, column=pct_col).value
        dec_is_100 = False
        if dec_val is not None:
            try:
                dec_is_100 = abs(float(dec_val) - 1.0) <= 0.005
            except (ValueError, TypeError):
                pass

        if monotonic_violations == 0 and dec_is_100:
            print(f"PASS: Component 4b — Monotonically increasing and December=100% (0.10 pts)")
            total_score += 0.10
        elif dec_is_100:
            print(f"PARTIAL: Component 4b — December=100% but not monotonically increasing (0.05 pts)")
            total_score += 0.05
        elif monotonic_violations == 0:
            print(f"PARTIAL: Component 4b — Monotonically increasing but December != 100% (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4b — Not monotonically increasing and December != 100%")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Grand Total row present with correct values (0.15 points)
    try:
        gt_row = None
        for r in range(data_start, pivot_ws.max_row + 1):
            val = pivot_ws.cell(row=r, column=month_col).value
            if val and isinstance(val, str) and 'total' in val.lower():
                gt_row = r
                break

        if gt_row is not None:
                gt_sales = pivot_ws.cell(row=gt_row, column=sales_col).value
                gt_pct = pivot_ws.cell(row=gt_row, column=pct_col).value if pct_col else None

                sales_ok = (gt_sales is not None and
                            abs(float(gt_sales) - EXPECTED_GRAND_TOTAL) <= 1.0)
                pct_ok = (gt_pct is not None and
                          abs(float(gt_pct) - 1.0) <= 0.005)

                if sales_ok and pct_ok:
                    print(f"PASS: Component 5 — Grand Total row: sales={gt_sales}, pct={gt_pct} (0.15 pts)")
                    total_score += 0.15
                elif sales_ok:
                    print(f"PARTIAL: Component 5 — Grand Total sales correct but % wrong (0.10 pts)")
                    total_score += 0.10
                elif pct_ok:
                    print(f"PARTIAL: Component 5 — Grand Total % correct but sales wrong (0.05 pts)")
                    total_score += 0.05
                else:
                    print(f"FAIL: Component 5 — Grand Total values incorrect: sales={gt_sales}, pct={gt_pct}")

        if gt_row is None:
            print(f"FAIL: Component 5 — No Grand Total row found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'

persist_app_state("libreoffice_calc")

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
