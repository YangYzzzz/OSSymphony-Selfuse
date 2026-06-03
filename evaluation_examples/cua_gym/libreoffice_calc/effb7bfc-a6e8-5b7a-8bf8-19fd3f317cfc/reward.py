"""
reward.py for task pdf_cross_038

Verifies ~/Documents/contacts.ods contains:
  1. File existence
  2. Correct number of rows (20 data rows + 1 header = 21 total)
  3. Correct column headers
  4. All 20 employees present (by email — unique identifier)
  5. Sort order: by department (A-Z) then by last name within dept (A-Z)
  6. Engineering rows (6) have light blue background
  7. Marketing rows (4) have light green background

Scoring (total = 1.0):
  - Component 1: File exists                    (0.05)
  - Component 2: Row count = 20                 (0.10)
  - Component 3: Column headers correct         (0.10)
  - Component 4: All 20 employees present       (0.25)
  - Component 5: Sort order correct             (0.20)
  - Component 6: Engineering rows light blue    (0.15)
  - Component 7: Marketing rows light green     (0.15)
"""

import os
import sys

ODS_PATH = "/home/user/Documents/contacts.ods"

# Canonical employee data (same as in setup/golden scripts)
EMPLOYEES = [
    ("James Wilson",      "Sales",       "james.wilson@company.com",       "555-201-0011"),
    ("Rachel Kim",        "Engineering", "rachel.kim@company.com",         "555-301-0021"),
    ("Tom Nguyen",        "Marketing",   "tom.nguyen@company.com",         "555-401-0031"),
    ("Linda Chen",        "HR",          "linda.chen@company.com",         "555-501-0041"),
    ("Carlos Ruiz",       "Engineering", "carlos.ruiz@company.com",        "555-301-0022"),
    ("Patricia Moore",    "Operations",  "patricia.moore@company.com",     "555-601-0051"),
    ("Kevin Zhang",       "Engineering", "kevin.zhang@company.com",        "555-301-0023"),
    ("Sandra Davis",      "Sales",       "sandra.davis@company.com",       "555-201-0012"),
    ("Ahmed Hassan",      "Marketing",   "ahmed.hassan@company.com",       "555-401-0032"),
    ("Megan Taylor",      "Engineering", "megan.taylor@company.com",       "555-301-0024"),
    ("Robert Jackson",    "HR",          "robert.jackson@company.com",     "555-501-0042"),
    ("Yuki Tanaka",       "Sales",       "yuki.tanaka@company.com",        "555-201-0013"),
    ("Elena Petrov",      "Engineering", "elena.petrov@company.com",       "555-301-0025"),
    ("Marcus Johnson",    "Operations",  "marcus.johnson@company.com",     "555-601-0052"),
    ("Diana Foster",      "Marketing",   "diana.foster@company.com",       "555-401-0033"),
    ("Chris Anderson",    "Sales",       "chris.anderson@company.com",     "555-201-0014"),
    ("Aisha Patel",       "Engineering", "aisha.patel@company.com",        "555-301-0026"),
    ("George Brown",      "Operations",  "george.brown@company.com",       "555-601-0053"),
    ("Natalie White",     "HR",          "natalie.white@company.com",      "555-501-0043"),
    ("Victor Cruz",       "Marketing",   "victor.cruz@company.com",        "555-401-0034"),
]

ENGINEERING_EMAILS = {e[2] for e in EMPLOYEES if e[1] == "Engineering"}
MARKETING_EMAILS   = {e[2] for e in EMPLOYEES if e[1] == "Marketing"}
ALL_EMAILS         = {e[2] for e in EMPLOYEES}

# Expected sort key: (dept.lower(), lastname.lower())
def _last_name(full: str) -> str:
    parts = (full or "").strip().split()
    return parts[-1].lower() if parts else ""

EXPECTED_ORDER = sorted(
    EMPLOYEES,
    key=lambda e: (e[1].lower(), _last_name(e[0])),
)

# Light blue / light green ARGB patterns we accept
# Allow for slight variations in how LibreOffice stores colors
BLUE_PATTERNS  = {"FFADD8E6", "ADD8E6"}   # light blue
GREEN_PATTERNS = {"FF90EE90", "90EE90"}   # light green

# Also allow hex values close to standard light-blue / light-green
# (LibreOffice may round RGB slightly)
def _is_light_blue(argb: str) -> bool:
    if not argb:
        return False
    argb = argb.upper().strip()
    if argb in BLUE_PATTERNS:
        return True
    # Strip leading FF if 8-char
    rgb = argb[-6:] if len(argb) >= 6 else argb
    try:
        r = int(rgb[0:2], 16)
        g = int(rgb[2:4], 16)
        b = int(rgb[4:6], 16)
        # Light blue: high blue, moderate green, lower red, blue dominates
        return r < 220 and b > 180 and g > 140 and b >= r
    except Exception:
        return False


def _is_light_green(argb: str) -> bool:
    if not argb:
        return False
    argb = argb.upper().strip()
    if argb in GREEN_PATTERNS:
        return True
    rgb = argb[-6:] if len(argb) >= 6 else argb
    try:
        r = int(rgb[0:2], 16)
        g = int(rgb[2:4], 16)
        b = int(rgb[4:6], 16)
        # Light green: green dominant, high green, lower red and blue
        return g > 180 and g > r and g > b
    except Exception:
        return False


def _get_fill_color(cell) -> str:
    """Return ARGB string of cell background, or empty string."""
    try:
        fill = cell.fill
        if fill and fill.fgColor:
            rgb = fill.fgColor.rgb
            if rgb and rgb not in ("00000000", "FFFFFFFF", "00FFFFFF"):
                return rgb.upper()
    except Exception:
        pass
    return ""


def verify_task():
    total_score = 0.0

    # ------------------------------------------------------------------
    # Component 1: File exists (0.05)
    # ------------------------------------------------------------------
    if os.path.exists(ODS_PATH):
        print(f"PASS: Component 1 — contacts.ods exists (0.05 pts)")
        total_score += 0.05
    else:
        print(f"FAIL: Component 1 — contacts.ods not found at {ODS_PATH}")
        print(f"REWARD: {round(total_score, 2)}")
        return total_score

    # Load workbook
    try:
        import openpyxl
        wb = openpyxl.load_workbook(ODS_PATH)
        ws = wb.active
    except Exception as e:
        print(f"FAIL: Could not open contacts.ods: {e}")
        print(f"REWARD: {round(total_score, 2)}")
        return total_score

    # ------------------------------------------------------------------
    # Component 2: Row count = 20 data rows (0.10)
    # ------------------------------------------------------------------
    # Find actual data rows (skip header at row 1)
    data_rows = []
    for r in range(2, ws.max_row + 1):
        name_val = ws.cell(row=r, column=1).value
        if name_val is not None and str(name_val).strip():
            data_rows.append(r)

    actual_count = len(data_rows)
    if actual_count == 20:
        print(f"PASS: Component 2 — 20 data rows found (0.10 pts)")
        total_score += 0.10
    else:
        print(f"FAIL: Component 2 — expected 20 data rows, found {actual_count}")

    # ------------------------------------------------------------------
    # Component 3: Column headers (0.10)
    # ------------------------------------------------------------------
    expected_headers = ["Name", "Department", "Email", "Phone"]
    actual_headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
    actual_headers_norm = [str(h).strip().lower() if h else "" for h in actual_headers]
    expected_headers_norm = [h.lower() for h in expected_headers]
    if actual_headers_norm == expected_headers_norm:
        print(f"PASS: Component 3 — column headers correct (0.10 pts)")
        total_score += 0.10
    else:
        print(f"FAIL: Component 3 — headers: expected {expected_headers}, got {actual_headers}")

    # ------------------------------------------------------------------
    # Component 4: All 20 employees present by email (0.25)
    # ------------------------------------------------------------------
    found_emails = set()
    for r in data_rows:
        email_val = ws.cell(row=r, column=3).value
        if email_val:
            found_emails.add(str(email_val).strip().lower())

    expected_emails_lower = {e.lower() for e in ALL_EMAILS}
    matched = expected_emails_lower & found_emails
    missing = expected_emails_lower - found_emails

    if len(matched) == 20:
        print(f"PASS: Component 4 — all 20 employees present (0.25 pts)")
        total_score += 0.25
    elif len(matched) >= 15:
        partial = round(0.25 * len(matched) / 20, 3)
        print(f"PARTIAL: Component 4 — {len(matched)}/20 employees found ({partial} pts)")
        total_score += partial
    else:
        print(f"FAIL: Component 4 — only {len(matched)}/20 employees found. Missing: {missing}")

    # ------------------------------------------------------------------
    # Component 5: Sort order — by dept then last name (0.20)
    # ------------------------------------------------------------------
    row_data = []
    for r in data_rows:
        name = ws.cell(row=r, column=1).value or ""
        dept = ws.cell(row=r, column=2).value or ""
        row_data.append((str(dept).strip(), str(name).strip()))

    sort_violations = 0
    for i in range(len(row_data) - 1):
        dept_i, name_i = row_data[i]
        dept_j, name_j = row_data[i + 1]
        key_i = (dept_i.lower(), _last_name(name_i))
        key_j = (dept_j.lower(), _last_name(name_j))
        if key_i > key_j:
            sort_violations += 1
            print(f"  sort violation at rows {data_rows[i]}/{data_rows[i+1]}: "
                  f"{dept_i}/{name_i} > {dept_j}/{name_j}")

    if sort_violations == 0:
        print(f"PASS: Component 5 — sort order correct (dept A-Z, last name A-Z) (0.20 pts)")
        total_score += 0.20
    elif sort_violations <= 2:
        partial = round(0.20 * (1 - sort_violations / max(len(row_data), 1)), 3)
        print(f"PARTIAL: Component 5 — {sort_violations} sort violations ({partial} pts)")
        total_score += partial
    else:
        print(f"FAIL: Component 5 — {sort_violations} sort violations")

    # ------------------------------------------------------------------
    # Component 6: Engineering rows have light blue background (0.15)
    # ------------------------------------------------------------------
    # Build email→row mapping
    email_to_row = {}
    for r in data_rows:
        email_val = ws.cell(row=r, column=3).value
        if email_val:
            email_to_row[str(email_val).strip().lower()] = r

    eng_correct = 0
    eng_total = 0
    for email in ENGINEERING_EMAILS:
        row_num = email_to_row.get(email.lower())
        if row_num is None:
            continue
        eng_total += 1
        color = _get_fill_color(ws.cell(row=row_num, column=1))
        if _is_light_blue(color):
            eng_correct += 1
        else:
            print(f"  Engineering row {row_num} color: '{color}' (expected light blue)")

    if eng_total == 0:
        print(f"FAIL: Component 6 — no Engineering rows found to check")
    elif eng_correct == eng_total:
        print(f"PASS: Component 6 — all {eng_total} Engineering rows have light blue (0.15 pts)")
        total_score += 0.15
    elif eng_correct > 0:
        partial = round(0.15 * eng_correct / eng_total, 3)
        print(f"PARTIAL: Component 6 — {eng_correct}/{eng_total} Engineering rows light blue ({partial} pts)")
        total_score += partial
    else:
        print(f"FAIL: Component 6 — 0/{eng_total} Engineering rows have light blue background")

    # ------------------------------------------------------------------
    # Component 7: Marketing rows have light green background (0.15)
    # ------------------------------------------------------------------
    mkt_correct = 0
    mkt_total = 0
    for email in MARKETING_EMAILS:
        row_num = email_to_row.get(email.lower())
        if row_num is None:
            continue
        mkt_total += 1
        color = _get_fill_color(ws.cell(row=row_num, column=1))
        if _is_light_green(color):
            mkt_correct += 1
        else:
            print(f"  Marketing row {row_num} color: '{color}' (expected light green)")

    if mkt_total == 0:
        print(f"FAIL: Component 7 — no Marketing rows found to check")
    elif mkt_correct == mkt_total:
        print(f"PASS: Component 7 — all {mkt_total} Marketing rows have light green (0.15 pts)")
        total_score += 0.15
    elif mkt_correct > 0:
        partial = round(0.15 * mkt_correct / mkt_total, 3)
        print(f"PARTIAL: Component 7 — {mkt_correct}/{mkt_total} Marketing rows light green ({partial} pts)")
        total_score += partial
    else:
        print(f"FAIL: Component 7 — 0/{mkt_total} Marketing rows have light green background")

    # ------------------------------------------------------------------
    final_score = round(min(total_score, 1.0), 2)
    print(f"\nREWARD: {final_score}")
    return final_score


if __name__ == "__main__":
    verify_task()
