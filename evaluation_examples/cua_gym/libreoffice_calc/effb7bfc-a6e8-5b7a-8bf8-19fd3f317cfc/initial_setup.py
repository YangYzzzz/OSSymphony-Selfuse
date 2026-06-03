"""
initial_setup.py for task pdf_cross_038

Creates:
  - ~/Documents/staff_directory.pdf  (2-page PDF listing 20 employees across 5 departments)
  - ~/Documents/contacts.ods         (empty/stub ODS file to be filled in by agent)

Then opens the PDF and the ODS in LibreOffice for the agent to work on.
"""

import os
import shlex
import subprocess
import time

TASK_ID = "pdf_cross_038"
WORKDIR = "/home/user/Documents"
PDF_PATH = f"{WORKDIR}/staff_directory.pdf"
ODS_PATH = f"{WORKDIR}/contacts.ods"

# 20 employees across 5 departments (not sorted by dept/lastname — agent must sort)
EMPLOYEES = [
    # name, department, email, phone
    ("James Wilson",      "Sales",       "james.wilson@company.com",       "555-201-0011"),
    ("Rachel Kim",        "Engineering", "rachel.kim@company.com",         "555-301-0021"),
    ("Tom Nguyen",        "Marketing",   "tom.nguyen@company.com",         "555-401-0031"),
    ("Linda Chen",        "HR",          "linda.chen@company.com",         "555-501-0041"),
    ("Carlos Ruiz",       "Engineering", "carlos.ruiz@company.com",        "555-301-0022"),
    ("Patricia Moore",    "Operations",  "patricia.moore@company.com",     "555-601-0051"),
    ("Kevin Zhang",       "Engineering", "kevin.zhang@company.com",        "555-301-0023"),
    ("Sandra Davis",      "Sales",       "sandra.davis@company.com",       "555-201-0012"),
    ("Ahmed Hassan",      "Marketing",   "ahmed.hassan@company.com",       "555-401-0032"),
    ("Megan Taylor",      "Engineering", "megan.taylor@company.com",       "555-301-0024"),
    ("Robert Jackson",    "HR",          "robert.jackson@company.com",     "555-501-0042"),
    ("Yuki Tanaka",       "Sales",       "yuki.tanaka@company.com",        "555-201-0013"),
    ("Elena Petrov",      "Engineering", "elena.petrov@company.com",       "555-301-0025"),
    ("Marcus Johnson",    "Operations",  "marcus.johnson@company.com",     "555-601-0052"),
    ("Diana Foster",      "Marketing",   "diana.foster@company.com",       "555-401-0033"),
    ("Chris Anderson",    "Sales",       "chris.anderson@company.com",     "555-201-0014"),
    ("Aisha Patel",       "Engineering", "aisha.patel@company.com",        "555-301-0026"),
    ("George Brown",      "Operations",  "george.brown@company.com",       "555-601-0053"),
    ("Natalie White",     "HR",          "natalie.white@company.com",      "555-501-0043"),
    ("Victor Cruz",       "Marketing",   "victor.cruz@company.com",        "555-401-0034"),
]


def create_pdf():
    """Create the staff_directory.pdf with 20 employees across 2 pages."""
    try:
        import pymupdf  # PyMuPDF
        HAS_PYMUPDF = True
    except ImportError:
        HAS_PYMUPDF = False

    os.makedirs(WORKDIR, exist_ok=True)

    if HAS_PYMUPDF:
        _create_pdf_pymupdf()
    else:
        _create_pdf_fpdf2()


def _create_pdf_pymupdf():
    import pymupdf

    doc = pymupdf.open()
    # Page dimensions: A4
    W, H = 595, 842

    # Split employees over 2 pages (10 per page)
    for page_idx in range(2):
        page = doc.new_page(width=W, height=H)
        employees_on_page = EMPLOYEES[page_idx * 10: (page_idx + 1) * 10]

        # Title
        page.insert_text(
            (40, 50),
            "Acme Corporation — Staff Directory",
            fontsize=16,
            fontname="helv",
            color=(0, 0, 0),
        )
        page.insert_text(
            (40, 70),
            f"Page {page_idx + 1} of 2",
            fontsize=10,
            fontname="helv",
            color=(0.4, 0.4, 0.4),
        )

        # Column headers
        headers = ["Name", "Department", "Email", "Phone"]
        col_x = [40, 160, 280, 450]
        y = 100
        for hdr, cx in zip(headers, col_x):
            page.insert_text((cx, y), hdr, fontsize=11, fontname="helv", color=(0, 0, 0.6))

        y = 120
        for emp in employees_on_page:
            name, dept, email, phone = emp
            vals = [name, dept, email, phone]
            for val, cx in zip(vals, col_x):
                page.insert_text((cx, y), val, fontsize=9, fontname="helv", color=(0, 0, 0))
            y += 22

        # Page marker (for reward verification)
        page.insert_text(
            (40, H - 30),
            f"STAFF_DIRECTORY_PAGE_{page_idx + 1}_MARKER",
            fontsize=6,
            fontname="helv",
            color=(0.9, 0.9, 0.9),
        )

    doc.save(PDF_PATH)
    doc.close()
    print(f"Created PDF: {PDF_PATH}")


def _create_pdf_fpdf2():
    from fpdf import FPDF

    class PDF(FPDF):
        pass

    pdf = PDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)

    for page_idx in range(2):
        pdf.add_page()
        employees_on_page = EMPLOYEES[page_idx * 10: (page_idx + 1) * 10]

        # Title
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 10, "Acme Corporation - Staff Directory", ln=True, align="C")
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 6, f"Page {page_idx + 1} of 2", ln=True, align="C")
        pdf.ln(4)

        # Headers
        pdf.set_font("Helvetica", "B", 11)
        col_widths = [55, 40, 65, 35]
        headers = ["Name", "Department", "Email", "Phone"]
        for hdr, w in zip(headers, col_widths):
            pdf.cell(w, 8, hdr, border=1)
        pdf.ln()

        # Rows
        pdf.set_font("Helvetica", "", 9)
        for emp in employees_on_page:
            name, dept, email, phone = emp
            for val, w in zip([name, dept, email, phone], col_widths):
                pdf.cell(w, 7, val, border=1)
            pdf.ln()

        # Page marker
        pdf.set_font("Helvetica", "", 5)
        pdf.set_text_color(230, 230, 230)
        pdf.cell(0, 4, f"STAFF_DIRECTORY_PAGE_{page_idx + 1}_MARKER", ln=True)
        pdf.set_text_color(0, 0, 0)

    pdf.output(PDF_PATH)
    print(f"Created PDF: {PDF_PATH}")


def create_stub_ods():
    """Create an empty contacts.ods stub with just headers (agent fills in the data)."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"

    # Headers only — agent will fill in the data from the PDF
    headers = ["Name", "Department", "Email", "Phone"]
    for c, hdr in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=hdr)

    # Save as xlsx first, then we use LibreOffice to convert — but for
    # simplicity we save with .ods extension directly; openpyxl doesn't truly
    # support ods, so we save as xlsx and rename. The agent will work with it
    # in LibreOffice which handles both.
    # Actually save a proper xlsx and name it contacts.ods so LibreOffice opens it.
    tmp_xlsx = ODS_PATH.replace(".ods", "_tmp.xlsx")
    wb.save(tmp_xlsx)

    # Use LibreOffice to convert to actual ODS format
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    result = subprocess.run(
        ["libreoffice", "--headless", "--convert-to", "ods", "--outdir", WORKDIR, tmp_xlsx],
        capture_output=True,
        text=True,
        env=env,
        timeout=60,
    )
    if os.path.exists(ODS_PATH):
        os.remove(tmp_xlsx)
        print(f"Created ODS: {ODS_PATH}")
    else:
        # Fallback: just rename the xlsx as ods
        os.rename(tmp_xlsx, ODS_PATH)
        print(f"Created stub ODS (xlsx-format): {ODS_PATH}")


def launch_gui(command: str, delay_sec: float = 1.5):
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def main():
    os.makedirs(WORKDIR, exist_ok=True)

    print("=== pdf_cross_038 initial_setup.py ===")
    print(f"Creating {PDF_PATH} ...")
    create_pdf()

    print(f"Creating {ODS_PATH} ...")
    create_stub_ods()

    # Verify files exist
    assert os.path.exists(PDF_PATH), f"PDF not created: {PDF_PATH}"
    assert os.path.exists(ODS_PATH), f"ODS not created: {ODS_PATH}"
    print(f"Files created successfully.")

    # Open the PDF in Evince (for reading) and the ODS in LibreOffice Calc
    launch_gui(f'evince "{PDF_PATH}"', delay_sec=1.5)
    launch_gui(f'libreoffice --calc "{ODS_PATH}"', delay_sec=2.0)

    print("GUI_READY: launched Evince (PDF) and LibreOffice Calc (ODS) with DISPLAY=:0")


if __name__ == "__main__":
    main()
