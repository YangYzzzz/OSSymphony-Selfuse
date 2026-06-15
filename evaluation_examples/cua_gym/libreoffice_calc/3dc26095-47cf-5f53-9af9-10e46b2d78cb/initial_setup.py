"""
Initial Setup: Deal velocity tracker - pre-task state with dates only
Task ID: calc_sales_066
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
from datetime import date

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_066'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Velocity'

    # --- Headers ---
    headers = [
        'Deal', 'Created', 'Qualified', 'Proposal',
        'Negotiation', 'Closed', 'Total Days',
        'Lead>Qual', 'Qual>Prop', 'Prop>Neg', 'Neg>Close'
    ]
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="000000")
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # --- Deal data (dates only, no formulas) ---
    deals = [
        ['D1', date(2025, 1, 5),  date(2025, 1, 20), date(2025, 2, 10),
               date(2025, 2, 25), date(2025, 3, 15)],
        ['D2', date(2025, 1, 12), date(2025, 2, 1),  date(2025, 3, 5),
               date(2025, 3, 20), date(2025, 4, 10)],
        ['D3', date(2025, 2, 1),  date(2025, 2, 15), date(2025, 3, 1),
               date(2025, 3, 10), date(2025, 3, 25)],
    ]

    date_format = 'yyyy-mm-dd'
    for r, row_data in enumerate(deals, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c + 1, value=row_data[c])
            cell.number_format = date_format

    # Columns G-K (7-11) are intentionally left empty for the task

    # --- Column widths ---
    ws.column_dimensions['A'].width = 10
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col_letter].width = 14
    for col_letter in ['G', 'H', 'I', 'J', 'K']:
        ws.column_dimensions[col_letter].width = 13

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
