"""
Reward Script: Deal velocity tracker with date-difference formulas, averages, and bottleneck identification
Task ID: calc_sales_066
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30) - Date-difference formulas in G2:K4
  Component 2 (0.25) - Average formulas in row 6 (G6:K6) with label in A6
  Component 3 (0.25) - Bottleneck row: A7 label + correct stage identified
  Component 4 (0.20) - Formula correctness: formulas reference correct cells
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_066'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that 'Velocity' sheet exists
    if 'Velocity' not in wb.sheetnames:
        print("CRITICAL: 'Velocity' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Velocity']

    # Component 1: Date-difference formulas in G2:K4 (0.30 points)
    # G=F-B, H=C-B, I=D-C, J=E-D, K=F-E for rows 2-4
    # These cells are None in the initial file, so this only passes on golden.
    try:
        formula_count = 0
        total_formula_cells = 15  # 5 columns x 3 rows

        expected_patterns = {
            'G': ('F', 'B'),  # G = F - B
            'H': ('C', 'B'),  # H = C - B
            'I': ('D', 'C'),  # I = D - C
            'J': ('E', 'D'),  # J = E - D
            'K': ('F', 'E'),  # K = F - E
        }

        for row in range(2, 5):  # rows 2, 3, 4
            for col_letter, (left, right) in expected_patterns.items():
                cell_val = ws[f'{col_letter}{row}'].value
                if cell_val is not None and isinstance(cell_val, str) and '=' in cell_val:
                    formula_count += 1

        if formula_count >= 12:
            print(f"PASS: Component 1 — {formula_count}/{total_formula_cells} date-difference formulas found (0.30 pts)")
            total_score += 0.30
        elif formula_count >= 6:
            partial = 0.15
            print(f"PARTIAL: Component 1 — {formula_count}/{total_formula_cells} formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {formula_count}/{total_formula_cells} date-difference formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Average formulas in row 6 with label (0.25 points)
    # G6:K6 should have AVERAGE formulas, A6 should say 'Average'
    try:
        avg_formula_count = 0
        has_avg_label = False

        # Check A6 label
        a6_val = ws['A6'].value
        has_avg_label = (a6_val is not None and 'average' in str(a6_val).lower())

        # Check G6:K6 for AVERAGE formulas
        for col_letter in ['G', 'H', 'I', 'J', 'K']:
            cell_val = ws[f'{col_letter}6'].value
            if cell_val is not None and isinstance(cell_val, str) and 'AVERAGE' in cell_val.upper():
                avg_formula_count += 1

        if has_avg_label and avg_formula_count >= 4:
            print(f"PASS: Component 2 — A6='{a6_val}', {avg_formula_count}/5 AVERAGE formulas in row 6 (0.25 pts)")
            total_score += 0.25
        elif avg_formula_count >= 3:
            partial = 0.15
            print(f"PARTIAL: Component 2 — label={'yes' if has_avg_label else 'no'}, {avg_formula_count}/5 AVERAGE formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — A6='{a6_val}', {avg_formula_count}/5 AVERAGE formulas in row 6")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Bottleneck identification in row 7 (0.25 points)
    # A7 should contain 'Bottleneck', and somewhere in row 7 should identify 'Qual>Prop'
    try:
        a7_val = ws['A7'].value
        has_bottleneck_label = a7_val is not None and 'bottleneck' in str(a7_val).lower()

        # Look for 'Qual>Prop' (or similar) in any cell of row 7
        bottleneck_found = ""  # empty string is falsy; set to cell value if found
        bottleneck_value = None
        for col in range(1, 12):  # columns A through K
            cell_val = ws.cell(row=7, column=col).value
            if cell_val is not None:
                cell_str = str(cell_val).strip()
                # Accept variations: 'Qual>Prop', 'Qual > Prop', 'Qualified>Proposal', etc.
                if 'qual' in cell_str.lower() and 'prop' in cell_str.lower():
                    bottleneck_found = cell_str  # non-empty string is truthy
                    bottleneck_value = cell_str
                    break

        if has_bottleneck_label and bottleneck_found:
            print(f"PASS: Component 3 — A7='{a7_val}', bottleneck='{bottleneck_value}' (0.25 pts)")
            total_score += 0.25
        elif has_bottleneck_label or bottleneck_found:
            partial = 0.10
            print(f"PARTIAL: Component 3 — label={'yes' if has_bottleneck_label else 'no'}, bottleneck={'yes' if bottleneck_found else 'no'} ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — A7='{a7_val}', no bottleneck stage identified in row 7")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Formula correctness — formulas reference correct cell pairs (0.20 points)
    # Verify that the formulas in G2:K4 actually reference the right columns
    try:
        correct_formulas = 0
        total_checks = 15

        expected_refs = {
            'G': ('F', 'B'),
            'H': ('C', 'B'),
            'I': ('D', 'C'),
            'J': ('E', 'D'),
            'K': ('F', 'E'),
        }

        for row in range(2, 5):
            for col_letter, (left_col, right_col) in expected_refs.items():
                cell_val = ws[f'{col_letter}{row}'].value
                if cell_val is not None and isinstance(cell_val, str):
                    formula_upper = cell_val.upper().replace(' ', '')
                    expected_left = f'{left_col}{row}'.upper()
                    expected_right = f'{right_col}{row}'.upper()
                    # Check that formula references both correct cells with subtraction
                    if expected_left in formula_upper and expected_right in formula_upper and '-' in formula_upper:
                        correct_formulas += 1

        if correct_formulas >= 12:
            print(f"PASS: Component 4 — {correct_formulas}/{total_checks} formulas reference correct cells (0.20 pts)")
            total_score += 0.20
        elif correct_formulas >= 6:
            partial = 0.10
            print(f"PARTIAL: Component 4 — {correct_formulas}/{total_checks} correct formula references ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {correct_formulas}/{total_checks} formulas have correct cell references")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
