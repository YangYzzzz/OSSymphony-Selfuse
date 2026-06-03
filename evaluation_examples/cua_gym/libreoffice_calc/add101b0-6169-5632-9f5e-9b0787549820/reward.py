"""
Reward Script: Volume discount VLOOKUP with approximate matching
Task ID: calc_sales_035
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): D2:D6 contain VLOOKUP formulas with TRUE (approximate match)
  Component 2 (0.3): E2:E6 contain final price formulas referencing Qty, UnitPrice, Discount
  Component 3 (0.4): VLOOKUP formulas reference the correct lookup table range and the
                      final price formula structure is mathematically correct (Qty*Price*(1-Disc))
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_035'

# Ground truth from task context
GROUND_TRUTH_D = {
    2: 0.00,   # Qty=5, min bracket 1 -> 0%
    3: 0.10,   # Qty=75, min bracket 50 -> 10%
    4: 0.15,   # Qty=200, min bracket 100 -> 15%
    5: 0.05,   # Qty=25, min bracket 10 -> 5%
    6: 0.20,   # Qty=600, min bracket 500 -> 20%
}

GROUND_TRUTH_E = {
    2: 500.0,    # 5 * 100 * (1 - 0.00) = 500
    3: 6750.0,   # 75 * 100 * (1 - 0.10) = 6750
    4: 17000.0,  # 200 * 100 * (1 - 0.15) = 17000
    5: 2375.0,   # 25 * 100 * (1 - 0.05) = 2375
    6: 48000.0,  # 600 * 100 * (1 - 0.20) = 48000
}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook for formula inspection
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Orders sheet exists
    if 'Orders' not in wb.sheetnames:
        print("CRITICAL: 'Orders' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Orders']

    # Component 1: D2:D6 contain VLOOKUP formulas with TRUE/approximate match (0.3 points)
    try:
        vlookup_count = 0
        for row in range(2, 7):
            cell_val = ws.cell(row=row, column=4).value  # D column
            if cell_val is not None and isinstance(cell_val, str):
                formula_upper = cell_val.upper().replace(" ", "")
                if '=VLOOKUP(' in formula_upper and ('TRUE' in formula_upper or ',1)' in formula_upper):
                    vlookup_count += 1
                    print(f"  D{row}: VLOOKUP with approximate match found: {cell_val}")
                else:
                    print(f"  D{row}: Formula found but missing VLOOKUP/TRUE: {cell_val}")
            else:
                print(f"  D{row}: No formula found, value={cell_val!r}")

        if vlookup_count == 5:
            print(f"PASS: Component 1 - All 5 VLOOKUP formulas with approximate match (0.3 pts)")
            total_score += 0.3
        elif vlookup_count >= 3:
            partial = round(0.3 * (vlookup_count / 5), 2)
            print(f"PARTIAL: Component 1 - {vlookup_count}/5 VLOOKUP formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - Only {vlookup_count}/5 VLOOKUP formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: E2:E6 contain final price formulas (0.3 points)
    try:
        formula_count = 0
        for row in range(2, 7):
            cell_val = ws.cell(row=row, column=5).value  # E column
            if cell_val is not None and isinstance(cell_val, str):
                formula_upper = cell_val.upper().replace(" ", "")
                # Must be a formula referencing B (Qty), C (UnitPrice), D (Discount)
                if cell_val.startswith('=') and 'B' in formula_upper and 'C' in formula_upper and 'D' in formula_upper:
                    formula_count += 1
                    print(f"  E{row}: Final price formula found: {cell_val}")
                else:
                    print(f"  E{row}: Formula found but doesn't reference B,C,D: {cell_val}")
            else:
                print(f"  E{row}: No formula found, value={cell_val!r}")

        if formula_count == 5:
            print(f"PASS: Component 2 - All 5 final price formulas present (0.3 pts)")
            total_score += 0.3
        elif formula_count >= 3:
            partial = round(0.3 * (formula_count / 5), 2)
            print(f"PARTIAL: Component 2 - {formula_count}/5 final price formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 - Only {formula_count}/5 final price formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Formula correctness validation (0.4 points)
    # Verify VLOOKUP references correct lookup table and final price formula
    # structure is mathematically correct
    try:
        correct_checks = 0
        total_checks = 10  # 5 D-column + 5 E-column

        for row in range(2, 7):
            # Check D column: VLOOKUP must reference VolumeDiscount sheet, column 2, approx match
            d_val = ws.cell(row=row, column=4).value
            if d_val is not None and isinstance(d_val, str):
                d_upper = d_val.upper().replace(" ", "")
                # Must lookup B{row} in VolumeDiscount A:B range, return column 2
                row_ref = f"B{row}"
                if (row_ref in d_upper and
                    'VOLUMEDISCOUNT' in d_upper and
                    ',2,' in d_upper):
                    correct_checks += 1
                    print(f"  D{row} structure: correct lookup of {row_ref} in VolumeDiscount col 2")
                else:
                    print(f"  D{row} structure: incorrect - missing proper references in {d_val}")
            else:
                print(f"  D{row} structure: no formula present")

            # Check E column: must compute Qty * UnitPrice * (1 - Discount)
            e_val = ws.cell(row=row, column=5).value
            if e_val is not None and isinstance(e_val, str):
                e_upper = e_val.upper().replace(" ", "")
                b_ref = f"B{row}"
                c_ref = f"C{row}"
                d_ref = f"D{row}"
                # Formula should reference B{row}, C{row}, D{row} and include (1-D{row}) pattern
                if (b_ref in e_upper and c_ref in e_upper and d_ref in e_upper and
                    ('1-' in e_upper or '(1-' in e_upper)):
                    correct_checks += 1
                    print(f"  E{row} structure: correct Qty*Price*(1-Disc) pattern")
                else:
                    print(f"  E{row} structure: incorrect pattern in {e_val}")
            else:
                print(f"  E{row} structure: no formula present")

        if correct_checks == 10:
            print(f"PASS: Component 3 - All 10 formula structures verified correct (0.4 pts)")
            total_score += 0.4
        elif correct_checks > 0:
            partial = round(0.4 * (correct_checks / 10), 2)
            print(f"PARTIAL: Component 3 - {correct_checks}/10 formula structures correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 - No formula structures matched expected patterns")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved LibreOffice state
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
