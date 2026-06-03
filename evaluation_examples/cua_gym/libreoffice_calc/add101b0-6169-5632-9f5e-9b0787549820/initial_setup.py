"""
Initial Setup: Create a price list with volume discount tiers and orders table.
Task ID: calc_sales_035
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_035'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: VolumeDiscount ---
    ws1 = wb.active
    ws1.title = 'VolumeDiscount'

    # Headers
    ws1['A1'] = 'Min Qty'
    ws1['B1'] = 'Discount %'

    # Style headers
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    for cell in [ws1['A1'], ws1['B1']]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.font = Font(bold=True, size=11, color="FFFFFF")

    # Discount tier data (sorted ascending by Min Qty for VLOOKUP approximate match)
    discount_data = [
        [1, 0.00],
        [10, 0.05],
        [50, 0.10],
        [100, 0.15],
        [500, 0.20],
    ]
    for r, row_data in enumerate(discount_data, 2):
        ws1.cell(row=r, column=1, value=row_data[0])
        ws1.cell(row=r, column=2, value=row_data[1])
        ws1.cell(row=r, column=2).number_format = '0%'

    # Column widths
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 14

    # --- Sheet 2: Orders ---
    ws2 = wb.create_sheet('Orders')

    # Headers
    headers = ['Order', 'Qty', 'Unit Price', 'Discount %', 'Final Price']
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Order data (D and E columns left EMPTY for the task)
    orders = [
        ['ORD1', 5, 100],
        ['ORD2', 75, 100],
        ['ORD3', 200, 100],
        ['ORD4', 25, 100],
        ['ORD5', 600, 100],
    ]
    for r, row_data in enumerate(orders, 2):
        ws2.cell(row=r, column=1, value=row_data[0])
        ws2.cell(row=r, column=2, value=row_data[1])
        ws2.cell(row=r, column=3, value=row_data[2])
        ws2.cell(row=r, column=3).number_format = '$#,##0.00'

    # D2:D6 and E2:E6 are intentionally left empty (task requires filling them)

    # Column widths
    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 14
    ws2.column_dimensions['D'].width = 14
    ws2.column_dimensions['E'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
