"""
Reward Script: Protect workbook structure with password 'WBStruct99'
Task ID: calc_adv_protect_workbook_014
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): Workbook structure lock is enabled (lockStructure == True)
  Component 2 (0.3): A workbook password hash is set (not None/empty)
  Component 3 (0.2): All 4 sheets remain unprotected at sheet level (cells editable)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'   # VM path — all reward scripts run on the VM
TASK_ID = 'calc_adv_protect_workbook_014'

EXPECTED_SHEETS = ['Overview', 'Q1', 'Q2', 'Q3']


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    The task requires:
      - Workbook structure protection enabled with password 'WBStruct99'
      - Sheet tabs cannot be inserted/deleted/moved/renamed
      - Cell data on all sheets remains editable (no sheet-level protection)

    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: workbook must have a security attribute
    security = getattr(wb, 'security', None)
    if security is None:
        print("INFO: No workbook security object found — workbook is unprotected.")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Component 1: Workbook structure lock is enabled (0.5 points)
    # lockStructure must be True to prevent insert/delete/move/rename of sheets
    # This FAILS on initial (lockStructure=False) and PASSES on golden (lockStructure=True)
    try:
        lock_structure = security.lockStructure
        if lock_structure is True:
            print(f"PASS: Component 1 — lockStructure is True (workbook structure is locked) (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — lockStructure is {repr(lock_structure)}, expected True")
    except Exception as e:
        print(f"ERROR: Component 1 — could not check lockStructure: {e}")

    # Component 2: A workbook password hash is present (0.3 points)
    # The password hash (workbookPassword) must be set — a non-None, non-empty value
    # indicates a password was applied. The hash '8ABE' corresponds to 'WBStruct99'.
    # This FAILS on initial (workbookPassword=None) and PASSES on golden (workbookPassword='8ABE')
    try:
        wb_password = security.workbookPassword
        if wb_password is not None and str(wb_password).strip() != '':
            print(f"PASS: Component 2 — workbookPassword hash is set (value: {wb_password}) (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — workbookPassword is {repr(wb_password)}, expected a non-empty hash")
    except Exception as e:
        print(f"ERROR: Component 2 — could not check workbookPassword: {e}")

    # Component 3: All 4 sheets remain unprotected at sheet level (0.2 points)
    # Task says cell data should still be editable — this means no sheet-level protection.
    # This check is awarded ONLY when workbook structure protection (Component 1) is active.
    # That compound condition ensures the initial file (lockStructure=False) scores 0.0 here.
    try:
        if total_score >= 0.5:  # only award if structure protection is actually active
            # Count sheets that are present and NOT sheet-protected
            protected_sheets = [
                name for name in EXPECTED_SHEETS
                if name in wb.sheetnames and wb[name].protection.sheet
            ]
            missing_sheets = [name for name in EXPECTED_SHEETS if name not in wb.sheetnames]
            sheets_ok = (len(missing_sheets) == 0 and len(protected_sheets) == 0)
            if missing_sheets:
                print(f"FAIL: Component 3 — missing expected sheets: {missing_sheets}")
            elif len(protected_sheets) > 0:
                print(f"FAIL: Component 3 — these sheets have sheet-level protection (cells not editable): "
                      f"{protected_sheets}")
            if sheets_ok:
                print(f"PASS: Component 3 — all 4 sheets ({', '.join(EXPECTED_SHEETS)}) have no "
                      f"sheet-level cell protection (cells remain editable) (0.2 pts)")
                total_score += 0.2
        else:
            print("SKIP: Component 3 — skipped because workbook structure lock (Component 1) is not active")
    except Exception as e:
        print(f"ERROR: Component 3 — could not check sheet protection: {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
