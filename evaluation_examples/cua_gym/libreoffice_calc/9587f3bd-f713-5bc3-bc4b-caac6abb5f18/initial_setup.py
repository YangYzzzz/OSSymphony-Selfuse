"""
Initial Setup: Protect workbook structure with password
Task ID: calc_adv_protect_workbook_014
Domain: libreoffice_calc

Creates a workbook with 4 sheets (Overview, Q1, Q2, Q3) containing realistic
quarterly business data. No workbook structure protection is active.
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_adv_protect_workbook_014'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Overview ----
    ws_overview = wb.active
    ws_overview.title = 'Overview'

    header_font = Font(bold=True, name='Calibri', size=11)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    white_font = Font(bold=True, name='Calibri', size=11, color='FFFFFFFF')

    # Title row
    ws_overview['A1'] = 'Northstar Analytics — Quarterly Performance Summary 2024'
    ws_overview['A1'].font = Font(bold=True, name='Calibri', size=14)
    ws_overview.merge_cells('A1:E1')
    ws_overview['A1'].alignment = Alignment(horizontal='center')
    ws_overview.row_dimensions[1].height = 30

    # Headers
    overview_headers = ['Quarter', 'Revenue ($)', 'Expenses ($)', 'Net Profit ($)', 'Growth (%)']
    for col, h in enumerate(overview_headers, 1):
        cell = ws_overview.cell(row=2, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    overview_data = [
        ['Q1 2024', 485230.00, 312480.00, 172750.00, 8.3],
        ['Q2 2024', 521800.00, 334200.00, 187600.00, 7.5],
        ['Q3 2024', 498750.00, 320100.00, 178650.00, 4.6],
        ['Q4 2024 (proj)', 560000.00, 355000.00, 205000.00, 12.2],
    ]
    for r, row_data in enumerate(overview_data, 3):
        for c, val in enumerate(row_data, 1):
            ws_overview.cell(row=r, column=c, value=val)

    # Totals row
    ws_overview.cell(row=7, column=1, value='Total / Avg').font = Font(bold=True)
    ws_overview.cell(row=7, column=2, value='=SUM(B3:B6)').font = Font(bold=True)
    ws_overview.cell(row=7, column=3, value='=SUM(C3:C6)').font = Font(bold=True)
    ws_overview.cell(row=7, column=4, value='=SUM(D3:D6)').font = Font(bold=True)
    ws_overview.cell(row=7, column=5, value='=AVERAGE(E3:E6)').font = Font(bold=True)

    ws_overview.column_dimensions['A'].width = 20
    ws_overview.column_dimensions['B'].width = 16
    ws_overview.column_dimensions['C'].width = 16
    ws_overview.column_dimensions['D'].width = 16
    ws_overview.column_dimensions['E'].width = 13

    # ---- Sheet 2: Q1 ----
    ws_q1 = wb.create_sheet('Q1')
    ws_q1['A1'] = 'Northstar Analytics — Q1 2024 Monthly Breakdown'
    ws_q1['A1'].font = Font(bold=True, name='Calibri', size=13)
    ws_q1.merge_cells('A1:F1')
    ws_q1['A1'].alignment = Alignment(horizontal='center')
    ws_q1.row_dimensions[1].height = 26

    q1_headers = ['Month', 'Product', 'Units Sold', 'Unit Price ($)', 'Revenue ($)', 'Region']
    for col, h in enumerate(q1_headers, 1):
        cell = ws_q1.cell(row=2, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    q1_data = [
        ['January', 'Enterprise Suite',   142, 890.00,  126380.00, 'North America'],
        ['January', 'Starter Pack',        89,  49.00,    4361.00, 'Europe'],
        ['January', 'Pro License',        215, 199.00,   42785.00, 'Asia Pacific'],
        ['February', 'Enterprise Suite',  138, 890.00,  122820.00, 'North America'],
        ['February', 'Starter Pack',       97,  49.00,    4753.00, 'Latin America'],
        ['February', 'Pro License',       228, 199.00,   45372.00, 'North America'],
        ['March', 'Enterprise Suite',     155, 890.00,  137950.00, 'North America'],
        ['March', 'Starter Pack',         104,  49.00,    5096.00, 'Europe'],
        ['March', 'Pro License',          245, 199.00,   48755.00, 'Asia Pacific'],
        ['March', 'Consulting',            12, 2500.00,  30000.00, 'North America'],
        ['March', 'Support Package',       66,  85.00,    5610.00, 'Europe'],
        ['March', 'Training Services',     18, 350.00,    6300.00, 'Asia Pacific'],
    ]
    for r, row_data in enumerate(q1_data, 3):
        for c, val in enumerate(row_data, 1):
            ws_q1.cell(row=r, column=c, value=val)

    ws_q1.cell(row=15, column=1, value='Q1 Total').font = Font(bold=True)
    ws_q1.cell(row=15, column=5, value='=SUM(E3:E14)').font = Font(bold=True)

    ws_q1.column_dimensions['A'].width = 12
    ws_q1.column_dimensions['B'].width = 22
    ws_q1.column_dimensions['C'].width = 13
    ws_q1.column_dimensions['D'].width = 16
    ws_q1.column_dimensions['E'].width = 14
    ws_q1.column_dimensions['F'].width = 18

    # ---- Sheet 3: Q2 ----
    ws_q2 = wb.create_sheet('Q2')
    ws_q2['A1'] = 'Northstar Analytics — Q2 2024 Monthly Breakdown'
    ws_q2['A1'].font = Font(bold=True, name='Calibri', size=13)
    ws_q2.merge_cells('A1:F1')
    ws_q2['A1'].alignment = Alignment(horizontal='center')
    ws_q2.row_dimensions[1].height = 26

    q2_headers = ['Month', 'Product', 'Units Sold', 'Unit Price ($)', 'Revenue ($)', 'Region']
    for col, h in enumerate(q2_headers, 1):
        cell = ws_q2.cell(row=2, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    q2_data = [
        ['April',  'Enterprise Suite',   161, 890.00,  143290.00, 'North America'],
        ['April',  'Starter Pack',       112,  49.00,    5488.00, 'Europe'],
        ['April',  'Pro License',        258, 199.00,   51342.00, 'Asia Pacific'],
        ['May',    'Enterprise Suite',   168, 890.00,  149520.00, 'North America'],
        ['May',    'Starter Pack',       121,  49.00,    5929.00, 'Europe'],
        ['May',    'Pro License',        270, 199.00,   53730.00, 'Latin America'],
        ['June',   'Enterprise Suite',   175, 890.00,  155750.00, 'North America'],
        ['June',   'Starter Pack',       130,  49.00,    6370.00, 'Europe'],
        ['June',   'Pro License',        283, 199.00,   56317.00, 'Asia Pacific'],
        ['June',   'Consulting',          15, 2500.00,  37500.00, 'North America'],
        ['June',   'Support Package',     78,  85.00,    6630.00, 'Europe'],
        ['June',   'Training Services',   22, 350.00,    7700.00, 'Asia Pacific'],
    ]
    for r, row_data in enumerate(q2_data, 3):
        for c, val in enumerate(row_data, 1):
            ws_q2.cell(row=r, column=c, value=val)

    ws_q2.cell(row=15, column=1, value='Q2 Total').font = Font(bold=True)
    ws_q2.cell(row=15, column=5, value='=SUM(E3:E14)').font = Font(bold=True)

    ws_q2.column_dimensions['A'].width = 12
    ws_q2.column_dimensions['B'].width = 22
    ws_q2.column_dimensions['C'].width = 13
    ws_q2.column_dimensions['D'].width = 16
    ws_q2.column_dimensions['E'].width = 14
    ws_q2.column_dimensions['F'].width = 18

    # ---- Sheet 4: Q3 ----
    ws_q3 = wb.create_sheet('Q3')
    ws_q3['A1'] = 'Northstar Analytics — Q3 2024 Monthly Breakdown'
    ws_q3['A1'].font = Font(bold=True, name='Calibri', size=13)
    ws_q3.merge_cells('A1:F1')
    ws_q3['A1'].alignment = Alignment(horizontal='center')
    ws_q3.row_dimensions[1].height = 26

    q3_headers = ['Month', 'Product', 'Units Sold', 'Unit Price ($)', 'Revenue ($)', 'Region']
    for col, h in enumerate(q3_headers, 1):
        cell = ws_q3.cell(row=2, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    q3_data = [
        ['July',      'Enterprise Suite',  152, 890.00,  135280.00, 'North America'],
        ['July',      'Starter Pack',       98,  49.00,    4802.00, 'Europe'],
        ['July',      'Pro License',       240, 199.00,   47760.00, 'Asia Pacific'],
        ['August',    'Enterprise Suite',  145, 890.00,  129050.00, 'North America'],
        ['August',    'Starter Pack',      103,  49.00,    5047.00, 'Latin America'],
        ['August',    'Pro License',       232, 199.00,   46168.00, 'North America'],
        ['September', 'Enterprise Suite',  160, 890.00,  142400.00, 'North America'],
        ['September', 'Starter Pack',      115,  49.00,    5635.00, 'Europe'],
        ['September', 'Pro License',       255, 199.00,   50745.00, 'Asia Pacific'],
        ['September', 'Consulting',         11, 2500.00,  27500.00, 'North America'],
        ['September', 'Support Package',    62,  85.00,    5270.00, 'Europe'],
        ['September', 'Training Services',  16, 350.00,    5600.00, 'Asia Pacific'],
    ]
    for r, row_data in enumerate(q3_data, 3):
        for c, val in enumerate(row_data, 1):
            ws_q3.cell(row=r, column=c, value=val)

    ws_q3.cell(row=15, column=1, value='Q3 Total').font = Font(bold=True)
    ws_q3.cell(row=15, column=5, value='=SUM(E3:E14)').font = Font(bold=True)

    ws_q3.column_dimensions['A'].width = 14
    ws_q3.column_dimensions['B'].width = 22
    ws_q3.column_dimensions['C'].width = 13
    ws_q3.column_dimensions['D'].width = 16
    ws_q3.column_dimensions['E'].width = 14
    ws_q3.column_dimensions['F'].width = 18

    # Confirm NO workbook protection is set (initial state)
    wb.security.lockStructure = False

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheets: {wb.sheetnames}')
    print(f'Workbook structure protection: {wb.security.lockStructure}')


create_initial()
