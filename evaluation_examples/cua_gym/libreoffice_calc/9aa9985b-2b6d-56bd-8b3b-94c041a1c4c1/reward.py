"""
Reward Script: Time tracking spreadsheet with weekly timesheet
Task ID: calc_wf_018
Domain: libreoffice_calc
Scoring:
  Component 1: Row total formulas (=SUM) in column G for each project (0.2 pts)
  Component 2: Daily total formulas in row 7 (0.2 pts)
  Component 3: Overtime formulas =MAX(0, daily_total - 8) in row 8 (0.15 pts)
  Component 4: Data validation on project code cells referencing master list (0.15 pts)
  Component 5: Conditional formatting on daily totals for overtime (0.1 pts)
  Component 6: Summary section with regular/OT hours and pay formulas (0.2 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_018'


def persist_app_state(domain):
    """Try to save any unsaved LibreOffice state via Ctrl+S."""
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print(f"PERSIST: ctrl+s sent for {domain}")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Timesheet' sheet exists
    if 'Timesheet' not in wb.sheetnames:
        print("CRITICAL: 'Timesheet' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Timesheet']

    # =========================================================================
    # Component 1: Row total formulas (=SUM) in column G for projects (0.2 pts)
    # Golden has G2:G6 with =SUM(Bx:Fx). Initial has no formulas in G column.
    # =========================================================================
    try:
        row_sum_count = 0
        for r in range(2, 7):  # rows 2-6 (5 projects)
            val = ws.cell(row=r, column=7).value  # column G
            if val is not None and isinstance(val, str) and 'SUM' in val.upper():
                row_sum_count += 1
        if row_sum_count >= 5:
            print(f"PASS: Component 1 — All 5 project row SUM formulas found in G2:G6 (0.2 pts)")
            total_score += 0.2
        elif row_sum_count >= 3:
            partial = 0.1
            print(f"PARTIAL: Component 1 — {row_sum_count}/5 project row SUM formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {row_sum_count}/5 project row SUM formulas found in G2:G6")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Daily total formulas in row 7, columns B-F (0.2 pts)
    # Golden has B7:F7 with =SUM(Bx:B6) etc. Initial row 7 has no formulas.
    # =========================================================================
    try:
        daily_total_count = 0
        for c in range(2, 7):  # columns B(2) through F(6)
            val = ws.cell(row=7, column=c).value
            if val is not None and isinstance(val, str) and 'SUM' in val.upper():
                daily_total_count += 1
        # Also check G7 for grand total
        g7_val = ws.cell(row=7, column=7).value
        has_g7_total = (g7_val is not None and isinstance(g7_val, str) and 'SUM' in str(g7_val).upper())

        if daily_total_count >= 5 and has_g7_total:
            print(f"PASS: Component 2 — All 5 daily total SUM formulas + grand total in row 7 (0.2 pts)")
            total_score += 0.2
        elif daily_total_count >= 3:
            partial = 0.1
            print(f"PARTIAL: Component 2 — {daily_total_count}/5 daily SUM formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {daily_total_count}/5 daily total SUM formulas in B7:F7")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Overtime formulas =MAX(0, daily_total - 8) in row 8 (0.15 pts)
    # Golden has B8:F8 with =MAX(0,Bx-8). Initial row 8 has no formulas.
    # =========================================================================
    try:
        ot_count = 0
        for c in range(2, 7):  # columns B through F
            val = ws.cell(row=8, column=c).value
            if val is not None and isinstance(val, str) and 'MAX' in val.upper():
                ot_count += 1
        # Also check G8 for total overtime
        g8_val = ws.cell(row=8, column=7).value
        has_g8_total = (g8_val is not None and isinstance(g8_val, str) and 'SUM' in str(g8_val).upper())

        if ot_count >= 5 and has_g8_total:
            print(f"PASS: Component 3 — All 5 overtime MAX formulas + OT total in row 8 (0.15 pts)")
            total_score += 0.15
        elif ot_count >= 3:
            partial = 0.08
            print(f"PARTIAL: Component 3 — {ot_count}/5 overtime formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {ot_count}/5 overtime MAX formulas in B8:F8")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: Data validation on project code cells (0.15 pts)
    # Golden has list validation on A2:A6 referencing Projects!$A$2:$A$9.
    # Initial has no data validation.
    # =========================================================================
    try:
        has_dv = False
        if ws.data_validations and ws.data_validations.dataValidation:
            for dv in ws.data_validations.dataValidation:
                if dv.type == 'list':
                    # Check that it references the Projects sheet
                    formula_str = str(dv.formula1) if dv.formula1 else ''
                    if 'Projects' in formula_str or 'projects' in formula_str.lower():
                        has_dv = True
                        break
                    # Also accept if it's a list-type DV applied to column A
                    sqref_str = str(dv.sqref) if dv.sqref else ''
                    if 'A' in sqref_str:
                        has_dv = True
                        break

        if has_dv:
            print(f"PASS: Component 4 — Data validation (list type) found on project code cells (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — No list data validation found referencing project master list")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # =========================================================================
    # Component 5: Conditional formatting on daily totals for overtime (0.1 pts)
    # Golden has conditional formatting on B7:F7. Initial has none.
    # =========================================================================
    try:
        has_cf = False
        cf_list = list(ws.conditional_formatting)
        if cf_list:
            for cf in cf_list:
                cf_range_str = str(cf)
                # Check if CF is applied to row 7 area (daily totals)
                if '7' in cf_range_str:
                    has_cf = True
                    break

        if has_cf:
            print(f"PASS: Component 5 — Conditional formatting found on daily totals row (0.1 pts)")
            total_score += 0.1
        else:
            print(f"FAIL: Component 5 — No conditional formatting found on daily totals")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # =========================================================================
    # Component 6: Summary section with regular/OT hours and pay formulas (0.2 pts)
    # Golden has rows 14-19 with Weekly Summary, formulas for regular hours,
    # OT hours, regular pay, OT pay, total pay.
    # Initial has NO summary section.
    # =========================================================================
    try:
        summary_score = 0.0

        # Check for summary section existence (look for labels in column A, rows 13-25)
        summary_found = False
        summary_start_row = None
        for r in range(10, 26):
            val = ws.cell(row=r, column=1).value
            if val and 'summary' in str(val).lower():
                summary_found = True
                summary_start_row = r
                break

        if not summary_found:
            print(f"FAIL: Component 6 — No summary section found")
        else:
            # Look for key labels and formulas in the summary section
            has_regular_hours = False
            has_ot_hours = False
            has_pay_formula = False

            for r in range(summary_start_row, min(summary_start_row + 10, ws.max_row + 1)):
                a_val = ws.cell(row=r, column=1).value
                b_val = ws.cell(row=r, column=2).value

                if a_val and 'regular' in str(a_val).lower() and 'hour' in str(a_val).lower():
                    # Check if B has a formula
                    if b_val is not None and isinstance(b_val, str) and '=' in str(b_val):
                        has_regular_hours = True

                if a_val and 'overtime' in str(a_val).lower() and 'hour' in str(a_val).lower():
                    if b_val is not None and isinstance(b_val, str) and '=' in str(b_val):
                        has_ot_hours = True

                if a_val and 'pay' in str(a_val).lower():
                    if b_val is not None and isinstance(b_val, str) and '=' in str(b_val):
                        has_pay_formula = True

            # Award partial credit
            sub_checks = [has_regular_hours, has_ot_hours, has_pay_formula]
            passed = sum(sub_checks)

            if passed == 3:
                summary_score = 0.2
                print(f"PASS: Component 6 — Summary section complete: regular hours, OT hours, pay formulas (0.2 pts)")
            elif passed >= 2:
                summary_score = 0.13
                print(f"PARTIAL: Component 6 — Summary section: {passed}/3 sub-checks passed ({summary_score} pts)")
            elif passed >= 1:
                summary_score = 0.07
                print(f"PARTIAL: Component 6 — Summary section: {passed}/3 sub-checks passed ({summary_score} pts)")
            else:
                print(f"FAIL: Component 6 — Summary section found but no formulas (regular_hrs={has_regular_hours}, ot_hrs={has_ot_hours}, pay={has_pay_formula})")

            total_score += summary_score
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
