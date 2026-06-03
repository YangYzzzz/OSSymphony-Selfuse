"""
Initial Setup: Time tracking weekly timesheet
Task ID: calc_wf_018
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_018'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Timesheet ---
    ws = wb.active
    ws.title = 'Timesheet'

    # Styling constants
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )
    label_font = Font(name='Calibri', size=11, bold=True)

    # Column widths
    ws.column_dimensions['A'].width = 18
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].width = 12

    # Row 1: Headers
    headers = ['Project Code', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Total']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Project codes and hours data (realistic)
    project_codes = ['PRJ-101', 'PRJ-205', 'PRJ-318', 'PRJ-422', 'PRJ-550']
    hours_data = [
        [3.0, 4.0, 2.5, 3.5, 2.0],   # PRJ-101
        [2.0, 1.5, 3.0, 2.0, 4.5],   # PRJ-205
        [1.5, 2.0, 1.0, 3.0, 1.5],   # PRJ-318
        [0.0, 1.0, 2.5, 0.0, 1.0],   # PRJ-422
        [2.5, 3.0, 0.0, 1.5, 2.0],   # PRJ-550
    ]

    data_font = Font(name='Calibri', size=11)
    data_align = Alignment(horizontal='center', vertical='center')
    project_font = Font(name='Calibri', size=11, bold=True)

    for r, (code, hours) in enumerate(zip(project_codes, hours_data), 2):
        # Project code in column A
        cell_a = ws.cell(row=r, column=1, value=code)
        cell_a.font = project_font
        cell_a.border = thin_border
        cell_a.alignment = Alignment(horizontal='left', vertical='center')

        # Hours in columns B-F
        for c, h_val in enumerate(hours, 2):
            cell = ws.cell(row=r, column=c, value=h_val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            cell.number_format = '0.0'

        # Column G (Total) - leave empty, task requires adding formulas
        cell_g = ws.cell(row=r, column=7)
        cell_g.border = thin_border

    # Row 7: Daily Total label (empty totals - task requires adding formulas)
    row_total = len(project_codes) + 2  # row 7
    ws.cell(row=row_total, column=1, value='Daily Total').font = label_font
    ws.cell(row=row_total, column=1).border = thin_border
    ws.cell(row=row_total, column=1).alignment = Alignment(horizontal='left', vertical='center')
    for c in range(2, 8):
        cell = ws.cell(row=row_total, column=c)
        cell.border = thin_border

    # Row 8: Overtime Hours label (empty - task requires adding overtime calc)
    row_ot = row_total + 1  # row 8
    ws.cell(row=row_ot, column=1, value='Overtime Hours').font = label_font
    ws.cell(row=row_ot, column=1).border = thin_border
    ws.cell(row=row_ot, column=1).alignment = Alignment(horizontal='left', vertical='center')
    for c in range(2, 8):
        cell = ws.cell(row=row_ot, column=c)
        cell.border = thin_border

    # Blank row 9

    # Row 10-11: Rate section
    ws.cell(row=10, column=1, value='Rate Information').font = Font(name='Calibri', size=12, bold=True, underline='single')
    ws.cell(row=11, column=1, value='Regular Rate ($/hr)')
    ws.cell(row=11, column=1).font = label_font
    ws.cell(row=11, column=2, value=45.00)
    ws.cell(row=11, column=2).number_format = '$#,##0.00'
    ws.cell(row=11, column=2).font = data_font
    ws.cell(row=12, column=1, value='Overtime Rate ($/hr)')
    ws.cell(row=12, column=1).font = label_font
    ws.cell(row=12, column=2, value=67.50)
    ws.cell(row=12, column=2).number_format = '$#,##0.00'
    ws.cell(row=12, column=2).font = data_font

    # NO summary section - task requires building it

    # --- Sheet 2: Projects (master list) ---
    ws2 = wb.create_sheet('Projects')
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 30

    # Header
    ws2.cell(row=1, column=1, value='Project Code').font = Font(name='Calibri', size=11, bold=True)
    ws2.cell(row=1, column=2, value='Project Name').font = Font(name='Calibri', size=11, bold=True)

    master_projects = [
        ('PRJ-101', 'Enterprise Platform Migration'),
        ('PRJ-205', 'Customer Analytics Dashboard'),
        ('PRJ-318', 'Mobile App Redesign'),
        ('PRJ-422', 'Cloud Infrastructure Setup'),
        ('PRJ-550', 'Data Pipeline Optimization'),
        ('PRJ-610', 'Security Audit & Compliance'),
        ('PRJ-715', 'API Gateway Integration'),
        ('PRJ-820', 'Machine Learning Pipeline'),
    ]
    for r, (code, name) in enumerate(master_projects, 2):
        ws2.cell(row=r, column=1, value=code)
        ws2.cell(row=r, column=2, value=name)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
