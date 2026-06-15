"""
Initial Setup: Create spreadsheet with 5 named ranges for named range deletion task
Task ID: calc_nrv_031
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.workbook.defined_name import DefinedName

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_031'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: Main data ---
    ws1 = wb.active
    ws1.title = "Sheet1"

    # Column A: Employee names (A1:A10 - referenced by 'test_range')
    ws1["A1"] = "Employee"
    employees = [
        "Sarah Chen", "Marcus Johnson", "Priya Patel", "David Kim",
        "Emma Rodriguez", "James Wilson", "Fatima Al-Hassan",
        "Lucas Fernandez", "Aisha Thompson"
    ]
    for i, name in enumerate(employees, 2):
        ws1.cell(row=i, column=1, value=name)

    # Column B: Department
    ws1["B1"] = "Department"
    depts = [
        "Engineering", "Marketing", "Finance", "Engineering",
        "HR", "Sales", "Engineering", "Marketing", "Finance"
    ]
    for i, dept in enumerate(depts, 2):
        ws1.cell(row=i, column=2, value=dept)

    # Column C: Active metrics (C1:C100 - referenced by 'ActiveData')
    ws1["C1"] = "Monthly Revenue"
    import random
    random.seed(42)
    for i in range(2, 102):
        ws1.cell(row=i, column=3, value=round(random.uniform(15000, 95000), 2))

    # Column D: Additional data
    ws1["D1"] = "Quarter"
    quarters = ["Q1", "Q2", "Q3", "Q4"] * 25
    for i, q in enumerate(quarters, 2):
        ws1.cell(row=i, column=4, value=q)

    # Column E: Summary stats (E1:E5 - referenced by 'old_summary')
    ws1["E1"] = "Summary Metric"
    ws1["E2"] = "Total Revenue"
    ws1["E3"] = "Average Revenue"
    ws1["E4"] = "Max Revenue"
    ws1["E5"] = "Min Revenue"

    # Column F: Summary values
    ws1["F1"] = "Value"
    ws1["F2"] = 4523000.50
    ws1["F3"] = 45230.01
    ws1["F4"] = 94850.75
    ws1["F5"] = 15120.30

    # Column G: Config values (G1:G3 - referenced by 'Config')
    ws1["G1"] = "Setting"
    ws1["G2"] = "FiscalYearStart"
    ws1["G3"] = "Currency"

    # Column H: Config values
    ws1["H1"] = "Value"
    ws1["H2"] = "January"
    ws1["H3"] = "USD"

    # --- Sheet2: Backup data ---
    ws2 = wb.create_sheet("Sheet2")

    # Column A: Item IDs
    ws2["A1"] = "Item ID"
    for i in range(2, 52):
        ws2.cell(row=i, column=1, value=f"ITM-{1000 + i}")

    # Column B: Backup values (B1:B50 - referenced by 'backup_data')
    ws2["B1"] = "Archived Amount"
    for i in range(2, 52):
        ws2.cell(row=i, column=2, value=round(random.uniform(500, 25000), 2))

    # Column C: Dates
    ws2["C1"] = "Archive Date"
    from datetime import date, timedelta
    base_date = date(2024, 1, 15)
    for i in range(2, 52):
        ws2.cell(row=i, column=3, value=base_date + timedelta(days=(i - 2) * 7))
        ws2.cell(row=i, column=3).number_format = 'yyyy-mm-dd'

    # --- Define Named Ranges ---
    # 'test_range' → Sheet1!$A$1:$A$10
    dn1 = DefinedName('test_range', attr_text="Sheet1!$A$1:$A$10")
    wb.defined_names.add(dn1)

    # 'backup_data' → Sheet2!$B$1:$B$50
    dn2 = DefinedName('backup_data', attr_text="Sheet2!$B$1:$B$50")
    wb.defined_names.add(dn2)

    # 'old_summary' → Sheet1!$E$1:$E$5
    dn3 = DefinedName('old_summary', attr_text="Sheet1!$E$1:$E$5")
    wb.defined_names.add(dn3)

    # 'ActiveData' → Sheet1!$C$1:$C$100
    dn4 = DefinedName('ActiveData', attr_text="Sheet1!$C$1:$C$100")
    wb.defined_names.add(dn4)

    # 'Config' → Sheet1!$G$1:$G$3
    dn5 = DefinedName('Config', attr_text="Sheet1!$G$1:$G$3")
    wb.defined_names.add(dn5)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Verify named ranges
    wb2 = openpyxl.load_workbook(OUTPUT)
    names = list(wb2.defined_names.keys()) if hasattr(wb2.defined_names, 'keys') else [dn.name for dn in wb2.defined_names.definedName]
    print(f'Named ranges in file: {names}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
