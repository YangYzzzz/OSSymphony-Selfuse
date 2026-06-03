"""
Initial Setup: Add profit column and Sheet2 concat label task
Task ID: osworld_calc_gross_profit_sheet2_concat_009
Domain: libreoffice_calc

Creates a spreadsheet with:
  - Sheet1: Store transaction data with Store Name, Revenue, Total Cost,
            and an empty Profit column (column D)
  - Sheet2: Empty Sheet2 (Sheet2!A1 must be empty initially)
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_gross_profit_sheet2_concat_009'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: Store Transaction Data ---
    ws1 = wb.active
    ws1.title = 'Sheet1'

    # Headers: Store Name, Revenue, Total Cost, Profit (empty)
    headers = ['Store Name', 'Revenue', 'Total Cost', 'Profit']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # Realistic store transaction data for "Green Leaf Market"
    store_name = 'Green Leaf Market'
    data = [
        [store_name, 48250.00, 31820.50],
        [store_name, 52640.75, 34190.00],
        [store_name, 39870.20, 26340.80],
        [store_name, 61380.00, 41250.00],
        [store_name, 45920.50, 30780.25],
        [store_name, 57340.00, 38620.00],
        [store_name, 43760.30, 29540.00],
        [store_name, 66850.00, 44710.50],
        [store_name, 51230.00, 34050.75],
        [store_name, 59410.00, 39870.00],
        [store_name, 47180.60, 31420.00],
        [store_name, 53920.00, 36140.30],
    ]

    for r, row_data in enumerate(data, 2):
        # Store Name (A), Revenue (B), Total Cost (C)
        ws1.cell(row=r, column=1, value=row_data[0])
        ws1.cell(row=r, column=2, value=row_data[1])
        ws1.cell(row=r, column=3, value=row_data[2])
        # Profit column D is intentionally left EMPTY (task requirement)

    # --- Sheet2: Empty (A1 must be empty for task to be valid) ---
    ws2 = wb.create_sheet('Sheet2')
    # Sheet2!A1 is intentionally left empty

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
