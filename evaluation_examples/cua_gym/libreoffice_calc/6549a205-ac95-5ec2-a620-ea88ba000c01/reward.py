"""
Reward Script: Add profit column and Sheet2 concat label with total profit
Task ID: osworld_calc_gross_profit_sheet2_concat_009
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.5 pts): Sheet1 column D (D2:D13) has profit formulas (=Bn-Cn)
  - Component 2 (0.3 pts): Sheet2 A1 contains a formula concatenating store name
                            with total profit using SUM of Sheet1!D column
  - Component 3 (0.2 pts): Sheet2 A1 formula uses TEXT with '#,##0.00' formatting
Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_gross_profit_sheet2_concat_009'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # --- Precondition: required sheets exist ---
    if 'Sheet1' not in wb.sheetnames:
        print("CRITICAL: Sheet1 not found in workbook")
        print("REWARD: 0.0")
        return 0.0
    if 'Sheet2' not in wb.sheetnames:
        print("CRITICAL: Sheet2 not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws1 = wb['Sheet1']
    ws2 = wb['Sheet2']

    # --- Component 1: Sheet1 column D has profit formulas (0.5 pts) ---
    # Each data row (rows 2 through max_row) should have =Bn-Cn in column D
    # Verifies the "add profit column computing revenue minus total cost" requirement
    try:
        # Find data rows in Sheet1: rows where B and C have numeric values
        data_rows = []
        for r in range(2, ws1.max_row + 1):
            b_val = ws1.cell(row=r, column=2).value
            c_val = ws1.cell(row=r, column=3).value
            if b_val is not None and c_val is not None:
                data_rows.append(r)

        if len(data_rows) == 0:
            print("FAIL: Component 1 — No data rows found in Sheet1 (B and C columns)")
        else:
            # Check that column D formulas exist for all data rows
            # Accept formulas like =Bn-Cn or =Bn - Cn (with spaces)
            profit_formula_rows = 0
            wrong_formula_rows = []
            missing_formula_rows = []

            for r in data_rows:
                d_val = ws1.cell(row=r, column=4).value
                if d_val is None:
                    missing_formula_rows.append(r)
                elif isinstance(d_val, str) and d_val.startswith('='):
                    # Normalize and check it matches =B{r}-C{r} pattern
                    normalized = d_val.replace(' ', '').upper()
                    expected_pattern = f'=B{r}-C{r}'
                    if normalized == expected_pattern:
                        profit_formula_rows += 1
                    else:
                        # Also check for valid alternative: =(Bn-Cn) or similar
                        # Accept any formula that subtracts column C from column B for same row
                        alt_pattern = re.compile(
                            r'^=\(?B' + str(r) + r'\s*-\s*C' + str(r) + r'\)?$',
                            re.IGNORECASE
                        )
                        if alt_pattern.match(d_val.replace(' ', '')):
                            profit_formula_rows += 1
                        else:
                            wrong_formula_rows.append((r, d_val))
                else:
                    # Might be a cached numeric value (data_only mode); treat as no formula
                    missing_formula_rows.append(r)

            total_data = len(data_rows)
            if profit_formula_rows == total_data and total_data > 0:
                print(f"PASS: Component 1 — All {total_data} data rows in Sheet1!D have profit formulas (=Bn-Cn) (0.5 pts)")
                total_score += 0.5
            elif profit_formula_rows > 0:
                partial = round(0.5 * profit_formula_rows / total_data, 4)
                print(f"PARTIAL: Component 1 — {profit_formula_rows}/{total_data} rows have correct profit formulas")
                if missing_formula_rows:
                    print(f"  Missing formulas in rows: {missing_formula_rows}")
                if wrong_formula_rows:
                    print(f"  Wrong formulas: {wrong_formula_rows}")
                total_score += partial
            else:
                print(f"FAIL: Component 1 — No profit formulas found in Sheet1!D")
                if missing_formula_rows:
                    print(f"  Missing in rows: {missing_formula_rows}")
                if wrong_formula_rows:
                    print(f"  Wrong formulas: {wrong_formula_rows}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # --- Component 2: Sheet2 A1 has concat formula with store name and total profit (0.3 pts) ---
    # Requirement: "in Sheet2 write a text label showing the combined store name and total profit"
    # Expected: =Sheet1!A2&" Total Profit: $"&TEXT(SUM(Sheet1!D2:D100),"#,##0.00")
    # or similar formula concatenating Sheet1!A2 with SUM of Sheet1 D column
    try:
        a1_val = ws2['A1'].value
        if a1_val is None:
            print("FAIL: Component 2 — Sheet2!A1 is empty; expected a formula combining store name and total profit")
        elif not isinstance(a1_val, str):
            print(f"FAIL: Component 2 — Sheet2!A1 is not a formula string, got: {repr(a1_val)}")
        elif not a1_val.startswith('='):
            print(f"FAIL: Component 2 — Sheet2!A1 is not a formula (does not start with '='): {repr(a1_val)}")
        else:
            # Normalize for checking
            norm = a1_val.replace(' ', '').upper()

            # Check for Sheet1!A2 reference (store name)
            has_store_ref = bool(re.search(r'SHEET1!A2', norm))

            # Check for SUM of Sheet1!D column reference
            has_sum_d = bool(re.search(r'SUM\(SHEET1!D\d+:D\d+\)', norm))

            # Check for concatenation operator (&)
            has_concat = '&' in norm

            if has_store_ref and has_sum_d and has_concat:
                print(f"PASS: Component 2 — Sheet2!A1 has formula concatenating store name (Sheet1!A2) with total profit SUM (0.3 pts)")
                print(f"  Formula: {a1_val}")
                total_score += 0.3
            else:
                missing = []
                if not has_store_ref:
                    missing.append("Sheet1!A2 reference (store name)")
                if not has_sum_d:
                    missing.append("SUM(Sheet1!D...) for total profit")
                if not has_concat:
                    missing.append("concatenation operator (&)")
                print(f"FAIL: Component 2 — Sheet2!A1 formula missing: {', '.join(missing)}")
                print(f"  Actual formula: {a1_val}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # --- Component 3: Sheet2 A1 formula uses TEXT function with '#,##0.00' format (0.2 pts) ---
    # Requirement: "total profit formatted to two decimal places"
    # Expected: TEXT(...,"#,##0.00") or similar with 2-decimal formatting
    try:
        a1_val = ws2['A1'].value
        if not isinstance(a1_val, str) or not a1_val.startswith('='):
            print("FAIL: Component 3 — Sheet2!A1 is not a formula; cannot verify TEXT formatting")
        else:
            norm_formula = a1_val.upper()
            has_text_func = 'TEXT(' in norm_formula

            # Check for 2-decimal format: #,##0.00 or 0.00 etc.
            has_two_decimal_format = bool(re.search(r'[#0],[#0]{2,}0\.00|0\.00', a1_val))

            if has_text_func and has_two_decimal_format:
                print(f"PASS: Component 3 — Sheet2!A1 uses TEXT() with 2-decimal format (0.2 pts)")
                total_score += 0.2
            elif has_text_func and not has_two_decimal_format:
                print(f"PARTIAL: Component 3 — Sheet2!A1 has TEXT() but format string not '#,##0.00' (0.1 pts)")
                print(f"  Formula: {a1_val}")
                total_score += 0.1
            else:
                print(f"FAIL: Component 3 — Sheet2!A1 formula does not use TEXT() function for formatting")
                print(f"  Formula: {a1_val}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
