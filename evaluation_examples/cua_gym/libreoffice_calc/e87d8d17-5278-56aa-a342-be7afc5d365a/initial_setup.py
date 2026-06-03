"""
Initial Setup: Fill Invoice ID column with sequential IDs
Task ID: osworld_calc_fill_sequence_numbers_004
Domain: libreoffice_calc

Creates a billing spreadsheet with 100 invoice rows.
Column A (Invoice ID) is intentionally empty -- the agent must fill it.
Columns B-F contain realistic billing data.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_fill_sequence_numbers_004'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ---- Sheet: Billing ----
    ws = wb.active
    ws.title = "Billing"

    # Headers (row 1)
    headers = ["Invoice ID", "Client Name", "Amount", "Due Date", "Status", "Category"]
    header_font = Font(bold=True, name="Calibri", size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name="Calibri", size=11, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # Realistic billing data (100 rows, rows 2-101)
    clients = [
        "Acme Corp", "BrightPath Solutions", "ClearView Analytics", "DeltaForce Systems",
        "EcoSmart Technologies", "FusionEdge Consulting", "GlobalReach Media", "HorizonTech Inc",
        "InnoBridge Partners", "JetStream Logistics", "KaleidoScope Design", "LightSpeed Networks",
        "MomentumPro Services", "NextLevel Ventures", "OmniCore Industries", "PeakPerform Group",
        "QuantumLeap Labs", "RedStar Marketing", "SkyBridge Finance", "TerraFirm Engineering",
        "UltraVision Studios", "VelocityPrime Corp", "WaveLength Digital", "XcelRate Solutions",
        "YieldBridge Capital",
    ]
    categories = ["Software", "Consulting", "Hardware", "Support", "Marketing", "Legal", "Design", "Training"]
    statuses = ["Paid", "Pending", "Overdue", "Processing"]

    import datetime, random
    random.seed(42)
    base_date = datetime.date(2024, 1, 1)

    data_rows = []
    for i in range(100):
        client = clients[i % len(clients)]
        amount = round(random.uniform(500.0, 25000.0), 2)
        days_offset = i * 3 + random.randint(0, 5)
        due_date = (base_date + datetime.timedelta(days=days_offset)).strftime("%Y-%m-%d")
        status = statuses[i % len(statuses)]
        category = categories[i % len(categories)]
        # Column A is intentionally left empty (no Invoice ID)
        data_rows.append(["", client, amount, due_date, status, category])

    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r_idx, row_data in enumerate(data_rows, 2):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val if val != "" else None)
            cell.border = data_border
            if c_idx == 3:  # Amount column
                cell.number_format = '#,##0.00'
            if c_idx == 4:  # Due Date column
                cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions["A"].width = 18  # Invoice ID
    ws.column_dimensions["B"].width = 26  # Client Name
    ws.column_dimensions["C"].width = 14  # Amount
    ws.column_dimensions["D"].width = 14  # Due Date
    ws.column_dimensions["E"].width = 14  # Status
    ws.column_dimensions["F"].width = 16  # Category

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f"Initial file created: {OUTPUT}")

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print("GUI_READY: launched LibreOffice Calc with DISPLAY=:0")


create_initial()
