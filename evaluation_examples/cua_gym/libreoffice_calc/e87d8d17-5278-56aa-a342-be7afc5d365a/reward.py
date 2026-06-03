"""
Reward Script: Fill Invoice ID column with sequential IDs INV-2024-001 to INV-2024-100
Task ID: osworld_calc_fill_sequence_numbers_004
Domain: libreoffice_calc
Scoring:
  Component 1: Column A rows 2-101 are non-empty (IDs filled in)         — 0.3 pts
  Component 2: All 100 IDs follow exact format 'INV-2024-XXX' (3-digit)  — 0.4 pts
  Component 3: IDs are sequential from 001 to 100 in correct row order   — 0.3 pts
  Total: 1.0
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_fill_sequence_numbers_004'

# Expected ID pattern
ID_PATTERN = re.compile(r'^INV-2024-(\d{3})$')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Initial state: Column A rows 2-101 are all empty (None).
    Golden state:  Column A rows 2-101 contain INV-2024-001 through INV-2024-100.
    """
    total_score = 0.0

    # Precondition: load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet named 'Billing' must exist
    if 'Billing' not in wb.sheetnames:
        print(f"CRITICAL: Sheet 'Billing' not found. Available: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Billing']

    # Collect column A values for rows 2–101 (data rows, not header)
    col_a_values = []
    for row in range(2, 102):
        col_a_values.append(ws.cell(row=row, column=1).value)

    # Component 1: All 100 Invoice ID cells in column A rows 2-101 are non-empty
    # This fails on initial_env (all None) and passes on golden_env (0.3 pts)
    try:
        filled_count = sum(1 for v in col_a_values if v is not None)
        if filled_count == 100:
            print(f"PASS: Component 1 — All 100 Invoice ID cells in column A are filled")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — Only {filled_count}/100 Invoice ID cells are filled (expected 100)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All 100 values follow the exact format 'INV-2024-NNN' (3-digit zero-padded)
    # This fails on initial_env (nothing filled) and passes on golden_env (0.4 pts)
    try:
        format_pass_count = 0
        format_fail_examples = []
        for i, val in enumerate(col_a_values):
            if val is None:
                format_fail_examples.append(f"Row {i+2}: None")
            elif not isinstance(val, str):
                format_fail_examples.append(f"Row {i+2}: non-string {repr(val)}")
            elif not ID_PATTERN.match(str(val)):
                format_fail_examples.append(f"Row {i+2}: format mismatch {repr(val)}")
            else:
                format_pass_count += 1

        if format_pass_count == 100:
            print(f"PASS: Component 2 — All 100 Invoice IDs match format 'INV-2024-NNN'")
            total_score += 0.4
        else:
            print(f"FAIL: Component 2 — Only {format_pass_count}/100 IDs match format 'INV-2024-NNN'")
            if format_fail_examples:
                print(f"  First failure examples: {format_fail_examples[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: IDs are sequential from INV-2024-001 to INV-2024-100 in correct row order
    # This fails on initial_env (no values) and passes on golden_env (0.3 pts)
    try:
        sequence_correct = 0
        sequence_errors = []
        for i, val in enumerate(col_a_values):
            expected_seq = i + 1  # rows 2..101 → sequences 1..100
            expected_id = f"INV-2024-{expected_seq:03d}"
            if str(val) == expected_id:
                sequence_correct += 1
            else:
                sequence_errors.append(f"Row {i+2}: expected '{expected_id}', found {repr(val)}")

        if sequence_correct == 100:
            print(f"PASS: Component 3 — All 100 IDs are in sequential order INV-2024-001 to INV-2024-100")
            total_score += 0.3
        else:
            print(f"FAIL: Component 3 — Only {sequence_correct}/100 IDs are in correct sequential position")
            if sequence_errors:
                print(f"  First error examples: {sequence_errors[:3]}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the VM environment
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
