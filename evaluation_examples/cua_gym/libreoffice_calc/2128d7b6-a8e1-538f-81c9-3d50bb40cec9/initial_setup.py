"""
Initial Setup: Create named ranges for salary data and use them in formulas for a compensation summary.
Task ID: calc_hr_050
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_050'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Data ---
    ws_data = wb.active
    ws_data.title = 'Data'

    # Headers
    ws_data['A1'] = 'Employee'
    ws_data['B1'] = 'Salary'

    # Style headers
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    for col in ['A', 'B']:
        cell = ws_data[f'{col}1']
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.font = Font(bold=True, size=11, color="FFFFFF")

    # Employee data - 10 rows with realistic names and salaries
    employees = [
        ('Sarah Chen', 95000),
        ('Marcus Johnson', 72000),
        ('Priya Patel', 118000),
        ('David Kim', 85000),
        ('Elena Rodriguez', 130000),
        ('James O\'Brien', 67000),
        ('Aisha Mohammed', 103000),
        ('Thomas Weber', 50000),
        ('Lisa Nakamura', 91000),
        ('Robert Gonzalez', 78000),
    ]

    for r, (name, salary) in enumerate(employees, 2):
        ws_data.cell(row=r, column=1, value=name)
        ws_data.cell(row=r, column=2, value=salary)
        ws_data.cell(row=r, column=2).number_format = '#,##0'

    # Set column widths
    ws_data.column_dimensions['A'].width = 22
    ws_data.column_dimensions['B'].width = 15

    # --- Sheet 2: Summary ---
    ws_summary = wb.create_sheet('Summary')

    # Headers
    ws_summary['A1'] = 'Metric'
    ws_summary['B1'] = 'Value'

    for col in ['A', 'B']:
        cell = ws_summary[f'{col}1']
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Metric labels - NO formulas or values in B column
    ws_summary['A2'] = 'Average'
    ws_summary['A3'] = 'Median'
    ws_summary['A4'] = 'Total'
    ws_summary['A5'] = 'Count'

    # B2:B5 intentionally left empty - task is to fill these with named range formulas

    # Set column widths
    ws_summary.column_dimensions['A'].width = 15
    ws_summary.column_dimensions['B'].width = 18

    # NO named ranges defined - that's part of the task

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
