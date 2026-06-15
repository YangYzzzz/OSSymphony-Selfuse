"""
Reward Script: Create named ranges for salary data and use them in formulas
Task ID: calc_hr_050
Domain: libreoffice_calc
Scoring:
  Component 1: Named range 'Salaries' exists for Data!$B$2:$B$11 (0.3 pts)
  Component 2: Summary B2 = =AVERAGE(Salaries) (0.2 pts)
  Component 3: Summary B3 = =MEDIAN(Salaries) (0.2 pts)
  Component 4: Summary B4 = =SUM(Salaries) (0.15 pts)
  Component 5: Summary B5 = =COUNT(Salaries) (0.15 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_050'


def check_formula(ws, coord, expected_func, expected_range_name):
    """
    Check if a cell contains a formula using the expected function and named range.
    Case-insensitive, whitespace-insensitive.
    Returns True if the formula matches.
    """
    val = ws[coord].value
    if not isinstance(val, str):
        return False, val
    normalized = val.upper().replace(" ", "")
    expected = f"={expected_func}({expected_range_name})".upper().replace(" ", "")
    return normalized == expected, val


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Data' and 'Summary' sheets must exist
    if 'Data' not in wb.sheetnames:
        print("FAIL: 'Data' sheet not found")
        print("REWARD: 0.0")
        return 0.0
    if 'Summary' not in wb.sheetnames:
        print("FAIL: 'Summary' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws_summary = wb['Summary']

    # Component 1: Named range 'Salaries' exists and references Data!$B$2:$B$11 (0.3 points)
    try:
        found_name = False
        correct_ref = False
        for name, dn in wb.defined_names.items():
            if name.lower() == 'salaries':
                found_name = True
                ref_val = dn.attr_text.upper().replace(" ", "")
                # Accept variations: Data!$B$2:$B$11, 'Data'!$B$2:$B$11
                expected_refs = [
                    "DATA!$B$2:$B$11",
                    "'DATA'!$B$2:$B$11",
                ]
                if ref_val in expected_refs:
                    correct_ref = True
                break

        if found_name and correct_ref:
            print(f"PASS: Component 1 - Named range 'Salaries' = Data!$B$2:$B$11 (0.3 pts)")
            total_score += 0.3
        elif found_name:
            print(f"FAIL: Component 1 - Named range 'Salaries' exists but references '{dn.attr_text}', expected Data!$B$2:$B$11")
        else:
            print(f"FAIL: Component 1 - Named range 'Salaries' not found")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Summary B2 = =AVERAGE(Salaries) (0.2 points)
    try:
        match, actual = check_formula(ws_summary, 'B2', 'AVERAGE', 'Salaries')
        if match:
            print(f"PASS: Component 2 - B2 = =AVERAGE(Salaries) (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 2 - Expected =AVERAGE(Salaries) in B2, found: {actual}")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Summary B3 = =MEDIAN(Salaries) (0.2 points)
    try:
        match, actual = check_formula(ws_summary, 'B3', 'MEDIAN', 'Salaries')
        if match:
            print(f"PASS: Component 3 - B3 = =MEDIAN(Salaries) (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 - Expected =MEDIAN(Salaries) in B3, found: {actual}")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Summary B4 = =SUM(Salaries) (0.15 points)
    try:
        match, actual = check_formula(ws_summary, 'B4', 'SUM', 'Salaries')
        if match:
            print(f"PASS: Component 4 - B4 = =SUM(Salaries) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 - Expected =SUM(Salaries) in B4, found: {actual}")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Summary B5 = =COUNT(Salaries) (0.15 points)
    try:
        match, actual = check_formula(ws_summary, 'B5', 'COUNT', 'Salaries')
        if match:
            print(f"PASS: Component 5 - B5 = =COUNT(Salaries) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 5 - Expected =COUNT(Salaries) in B5, found: {actual}")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
