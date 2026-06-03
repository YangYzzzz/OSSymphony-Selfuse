"""
Reward Script: Find & Replace PRD- codes with ITEM- codes
Task ID: calc_gg5_019
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): No PRD- codes remain in column A
  Component 2 (0.3): All 80 data codes match ITEM-XXXX format
  Component 3 (0.3): Header preserved and exactly 80 data rows present
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_019'


def persist_app_state(domain: str):
    """Try to save any unsaved GUI state."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Products' sheet must exist
    if 'Products' not in wb.sheetnames:
        print("CRITICAL: 'Products' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Products']

    # Collect all product code values from column A (data rows: 2 to max_row)
    codes = []
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val is not None:
            codes.append(str(val))

    prd_count = sum(1 for c in codes if c.startswith('PRD-'))
    item_pattern = re.compile(r'^ITEM-\d{4}$')
    item_count = sum(1 for c in codes if item_pattern.match(c))

    print(f"INFO: Found {len(codes)} non-empty codes in column A")
    print(f"INFO: PRD- codes: {prd_count}, ITEM-XXXX codes: {item_count}")

    # Component 1: No PRD- codes remain (0.4 points)
    # This FAILS on initial (80 PRD- codes) and PASSES on golden (0 PRD- codes)
    try:
        if prd_count == 0:
            print(f"PASS: Component 1 — No PRD- codes remain (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — {prd_count} PRD- codes still present")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All 80 data codes match ITEM-XXXX pattern (0.3 points)
    # This FAILS on initial (0 ITEM- codes) and PASSES on golden (80 ITEM- codes)
    try:
        if item_count == 80:
            print(f"PASS: Component 2 — All 80 codes match ITEM-XXXX format (0.3 pts)")
            total_score += 0.3
        elif item_count > 0:
            # Partial credit: proportion of codes correctly converted
            partial = 0.3 * (item_count / 80)
            print(f"PARTIAL: Component 2 — {item_count}/80 codes match ITEM-XXXX ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No ITEM-XXXX codes found (found {len(codes)} codes total)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Header preserved as 'Product Code' and exactly 80 data rows (0.3 points)
    # Sub-check A: header is 'Product Code' (precondition-style but combined with count check)
    # Sub-check B: exactly 80 non-empty data codes exist
    # This FAILS on initial because we gate it behind Component 1 passing (no PRD- codes)
    # Actually — we need this to independently fail on initial.
    # Redesign: Check that all codes are ITEM- AND total count is 80 AND header is preserved.
    # On initial: codes are PRD- so item_count=0, this fails.
    try:
        header_val = ws.cell(row=1, column=1).value
        header_ok = (header_val == 'Product Code')
        count_ok = (len(codes) == 80)
        all_item = (item_count == 80)

        if header_ok and count_ok and all_item:
            print(f"PASS: Component 3 — Header='Product Code', 80 data rows, all ITEM- format (0.3 pts)")
            total_score += 0.3
        else:
            reasons = []
            if not header_ok:
                reasons.append(f"header is '{header_val}' not 'Product Code'")
            if not count_ok:
                reasons.append(f"found {len(codes)} data rows, expected 80")
            if not all_item:
                reasons.append(f"only {item_count}/80 are ITEM-XXXX")
            print(f"FAIL: Component 3 — {'; '.join(reasons)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
