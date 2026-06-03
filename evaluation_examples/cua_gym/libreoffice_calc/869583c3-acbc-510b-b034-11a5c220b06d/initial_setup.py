"""
Initial Setup: Find & Replace product codes from PRD- to ITEM- prefix
Task ID: calc_gg5_019
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_019'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Products'

    # --- Headers ---
    headers = ['Product Code', 'Product Name', 'Category', 'Unit Price', 'Stock Qty', 'Supplier']
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Product data generation ---
    product_names = [
        'Wireless Bluetooth Headphones', 'USB-C Charging Cable', 'Ergonomic Office Chair',
        'Mechanical Keyboard RGB', 'Portable External SSD 1TB', 'Smart LED Desk Lamp',
        'Noise Cancelling Earbuds', 'Adjustable Monitor Stand', 'Webcam HD 1080p',
        'Laptop Cooling Pad', 'Wireless Mouse Ergonomic', 'HDMI to DisplayPort Adapter',
        'Surge Protector Power Strip', 'Desktop Microphone USB', 'Cable Management Kit',
        'Screen Cleaning Kit', 'Phone Stand Adjustable', 'Portable Charger 20000mAh',
        'Smart Power Plug WiFi', 'Document Scanner Portable', 'LED Ring Light 12 inch',
        'Thunderbolt Docking Station', 'Laptop Sleeve 15 inch', 'Wireless Charging Pad',
        'Bluetooth Speaker Portable', 'Network Switch 8-Port', 'UPS Battery Backup 600VA',
        'Graphics Tablet Drawing', 'Thermal Printer Label', 'Barcode Scanner Wireless',
        'Receipt Printer Thermal', 'Cash Drawer POS', 'Touch Screen Monitor 22 inch',
        'KVM Switch 2-Port', 'Ethernet Cable Cat6 50ft', 'WiFi Range Extender AC1200',
        'Smart Thermostat WiFi', 'Air Quality Monitor', 'Desk Organizer Bamboo',
        'Whiteboard Magnetic 48x36', 'Laser Pointer Presenter', 'ID Card Printer',
        'Shredder Cross Cut 12 Sheet', 'Laminator Machine A4', 'Binding Machine Comb',
        'Projector HD 1080p', 'Projection Screen 100 inch', 'Conference Phone Speaker',
        'Webcam Light Bar', 'USB Hub 7-Port Powered', 'Memory Card Reader USB3',
        'Flash Drive 128GB', 'External Hard Drive 4TB', 'NVMe Enclosure USB-C',
        'Deskpad XXL Leather', 'Footrest Ergonomic', 'Lumbar Support Pillow',
        'Anti Fatigue Mat Standing', 'Monitor Privacy Screen 27 inch', 'Blue Light Glasses',
        'Keyboard Wrist Rest Gel', 'Mouse Pad Extended XXL', 'Cable Clips Adhesive 50pk',
        'Velcro Cable Ties 100pk', 'Label Maker Portable', 'Tape Dispenser Heavy Duty',
        'Stapler Electric Desktop', 'Paper Trimmer Guillotine', 'Hole Punch 3-Ring',
        'File Cabinet 2-Drawer', 'Desk Fan USB Quiet', 'Space Heater Ceramic',
        'Humidifier Ultrasonic', 'Air Purifier HEPA Desktop', 'Noise Machine White',
        'Standing Desk Converter', 'Sit Stand Desk Electric 60', 'Monitor Arm Dual',
        'Laptop Stand Aluminum', 'Docking Station USB-C Triple', 'Presentation Clicker Wireless',
    ]

    categories = [
        'Audio', 'Cables & Adapters', 'Furniture', 'Input Devices', 'Storage',
        'Lighting', 'Audio', 'Furniture', 'Video', 'Cooling',
        'Input Devices', 'Cables & Adapters', 'Power', 'Audio', 'Accessories',
        'Cleaning', 'Accessories', 'Power', 'Smart Home', 'Scanners',
        'Lighting', 'Docking', 'Cases & Bags', 'Power', 'Audio',
        'Networking', 'Power', 'Input Devices', 'Printers', 'Scanners',
        'Printers', 'POS Equipment', 'Displays', 'Networking', 'Cables & Adapters',
        'Networking', 'Smart Home', 'Monitoring', 'Accessories', 'Office Supplies',
        'Presentation', 'Printers', 'Office Equipment', 'Office Equipment', 'Office Equipment',
        'Displays', 'Presentation', 'Audio', 'Lighting', 'Connectivity',
        'Connectivity', 'Storage', 'Storage', 'Storage', 'Accessories',
        'Furniture', 'Furniture', 'Furniture', 'Privacy', 'Eyewear',
        'Accessories', 'Accessories', 'Accessories', 'Accessories', 'Office Supplies',
        'Office Supplies', 'Office Equipment', 'Office Equipment', 'Office Equipment', 'Furniture',
        'Climate', 'Climate', 'Climate', 'Climate', 'Climate',
        'Furniture', 'Furniture', 'Furniture', 'Accessories', 'Docking', 'Presentation',
    ]

    suppliers = [
        'TechVenture Inc.', 'GlobalLink Supply', 'PremiumParts Co.', 'NovaTech Solutions',
        'Pinnacle Distribution', 'EastBridge Trading', 'SilverLine Industries', 'CoreTech Wholesale',
        'BrightStar Electronics', 'OmniSource Ltd.', 'Vertex Supply Chain', 'MeridianTech',
    ]

    # Generate 80 unique 4-digit product codes
    used_codes = set()
    product_codes = []
    while len(product_codes) < 80:
        code = random.randint(1, 9999)
        code_str = f'{code:04d}'
        if code_str not in used_codes:
            used_codes.add(code_str)
            product_codes.append(code_str)

    # Prices
    price_ranges = {
        'Audio': (29.99, 299.99), 'Cables & Adapters': (7.99, 49.99),
        'Furniture': (79.99, 599.99), 'Input Devices': (24.99, 179.99),
        'Storage': (19.99, 249.99), 'Lighting': (19.99, 89.99),
        'Video': (39.99, 159.99), 'Cooling': (14.99, 49.99),
        'Power': (19.99, 199.99), 'Accessories': (9.99, 69.99),
        'Cleaning': (7.99, 24.99), 'Smart Home': (24.99, 149.99),
        'Scanners': (49.99, 299.99), 'Docking': (89.99, 349.99),
        'Cases & Bags': (14.99, 59.99), 'Networking': (19.99, 149.99),
        'Monitoring': (49.99, 199.99), 'Office Supplies': (9.99, 79.99),
        'Presentation': (24.99, 499.99), 'Printers': (99.99, 499.99),
        'POS Equipment': (49.99, 299.99), 'Displays': (149.99, 599.99),
        'Privacy': (29.99, 79.99), 'Eyewear': (14.99, 49.99),
        'Connectivity': (12.99, 89.99), 'Office Equipment': (29.99, 299.99),
        'Climate': (24.99, 199.99),
    }

    for i in range(80):
        row = i + 2
        code = f'PRD-{product_codes[i]}'
        name = product_names[i % len(product_names)]
        cat = categories[i % len(categories)]
        low, high = price_ranges.get(cat, (9.99, 199.99))
        price = round(random.uniform(low, high), 2)
        stock = random.randint(0, 500)
        supplier = random.choice(suppliers)

        ws.cell(row=row, column=1, value=code)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=cat)
        price_cell = ws.cell(row=row, column=4, value=price)
        price_cell.number_format = '$#,##0.00'
        ws.cell(row=row, column=5, value=stock)
        ws.cell(row=row, column=6, value=supplier)

    # Column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 24

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
