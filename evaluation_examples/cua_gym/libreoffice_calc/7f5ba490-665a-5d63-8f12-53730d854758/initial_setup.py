"""
Initial Setup: Highlight rows where Priority is Critical AND Status is not Resolved
Task ID: calc_gcv_051
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_051'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Incident_Register"

    # --- Headers ---
    headers = ["Incident ID", "Title", "Category", "Priority", "Assigned To", "Status"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Data: 39 incident rows ---
    categories = ["Network", "Hardware", "Software", "Security", "Database", "Application", "Infrastructure"]
    priorities = ["Critical", "High", "Medium", "Low"]
    statuses = ["Open", "In Progress", "Resolved"]

    incidents = [
        ["INC-2025-001", "Production database server unresponsive", "Database", "Critical", "Sarah Chen", "Open"],
        ["INC-2025-002", "Email server latency exceeding 5s", "Network", "High", "Marcus Johnson", "In Progress"],
        ["INC-2025-003", "Firewall rule misconfiguration detected", "Security", "Critical", "Priya Sharma", "Open"],
        ["INC-2025-004", "Printer on Floor 3 jamming repeatedly", "Hardware", "Low", "James Wilson", "Resolved"],
        ["INC-2025-005", "CRM application login failures", "Application", "High", "Elena Rodriguez", "Open"],
        ["INC-2025-006", "Backup job failed for finance DB", "Database", "Critical", "David Kim", "In Progress"],
        ["INC-2025-007", "VPN connectivity drops during peak hours", "Network", "High", "Aisha Patel", "Open"],
        ["INC-2025-008", "Laptop battery swelling reported", "Hardware", "Medium", "Tom Bradley", "In Progress"],
        ["INC-2025-009", "Unauthorized access attempt on admin portal", "Security", "Critical", "Sarah Chen", "Resolved"],
        ["INC-2025-010", "SAP module crashing on report generation", "Software", "High", "Marcus Johnson", "Open"],
        ["INC-2025-011", "DNS resolution failures intermittent", "Network", "Critical", "Priya Sharma", "Open"],
        ["INC-2025-012", "Monitor flickering on workstation WS-042", "Hardware", "Low", "James Wilson", "Resolved"],
        ["INC-2025-013", "SSL certificate expiring in 7 days", "Security", "High", "Elena Rodriguez", "In Progress"],
        ["INC-2025-014", "Data warehouse ETL pipeline stalled", "Database", "Critical", "David Kim", "Open"],
        ["INC-2025-015", "Office 365 sync issues across department", "Application", "Medium", "Aisha Patel", "Open"],
        ["INC-2025-016", "Server room temperature alert triggered", "Infrastructure", "Critical", "Tom Bradley", "In Progress"],
        ["INC-2025-017", "Wireless AP on Building B floor 2 down", "Network", "Medium", "Sarah Chen", "Resolved"],
        ["INC-2025-018", "Keyboard replacement request batch", "Hardware", "Low", "Marcus Johnson", "Open"],
        ["INC-2025-019", "Ransomware signature detected in sandbox", "Security", "Critical", "Priya Sharma", "Open"],
        ["INC-2025-020", "Accounting software rounding errors", "Software", "High", "James Wilson", "In Progress"],
        ["INC-2025-021", "Load balancer failover not triggering", "Infrastructure", "Critical", "Elena Rodriguez", "Open"],
        ["INC-2025-022", "Projector bulb replacement conference room A", "Hardware", "Low", "David Kim", "Resolved"],
        ["INC-2025-023", "Customer portal session timeouts", "Application", "High", "Aisha Patel", "Open"],
        ["INC-2025-024", "Replication lag on read replica exceeding 30s", "Database", "Critical", "Tom Bradley", "In Progress"],
        ["INC-2025-025", "Windows update causing BSOD on dev machines", "Software", "High", "Sarah Chen", "Open"],
        ["INC-2025-026", "Badge reader malfunction at main entrance", "Hardware", "Medium", "Marcus Johnson", "In Progress"],
        ["INC-2025-027", "DDoS mitigation rules need updating", "Security", "Critical", "Priya Sharma", "Resolved"],
        ["INC-2025-028", "Cloud storage quota exceeded for marketing", "Infrastructure", "Medium", "James Wilson", "Open"],
        ["INC-2025-029", "Mobile app crash on Android 14 devices", "Application", "High", "Elena Rodriguez", "In Progress"],
        ["INC-2025-030", "Core switch firmware vulnerability", "Network", "Critical", "David Kim", "Open"],
        ["INC-2025-031", "Docking station USB-C port failures", "Hardware", "Medium", "Aisha Patel", "Open"],
        ["INC-2025-032", "Jenkins build pipeline timeout errors", "Software", "High", "Tom Bradley", "Resolved"],
        ["INC-2025-033", "Privilege escalation vulnerability in LDAP", "Security", "Critical", "Sarah Chen", "Open"],
        ["INC-2025-034", "PostgreSQL connection pool exhaustion", "Database", "High", "Marcus Johnson", "In Progress"],
        ["INC-2025-035", "UPS battery replacement overdue in DC-East", "Infrastructure", "Critical", "Priya Sharma", "In Progress"],
        ["INC-2025-036", "Zoom integration broken after update", "Application", "Medium", "James Wilson", "Resolved"],
        ["INC-2025-037", "VLAN segmentation audit findings", "Network", "High", "Elena Rodriguez", "Open"],
        ["INC-2025-038", "SSD failure predicted on server SRV-017", "Hardware", "Critical", "David Kim", "Open"],
        ["INC-2025-039", "API gateway rate limiting misconfigured", "Software", "High", "Aisha Patel", "In Progress"],
    ]

    for r, row_data in enumerate(incidents, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 1:
                cell.alignment = Alignment(horizontal="center")

    # Set column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 14

    # Freeze header row
    ws.freeze_panes = "A2"

    # NO conditional formatting in initial state
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
