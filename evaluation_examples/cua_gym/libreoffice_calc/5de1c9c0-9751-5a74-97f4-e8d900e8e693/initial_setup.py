"""
Initial Setup: Analyze survey results in LibreOffice Calc
Task ID: calc_grs_010
Domain: libreoffice_calc

Creates a workbook with raw survey data from 50 respondents:
- 5 Likert-scale questions (Q1-Q5, rated 1-5)
- 2 multiple choice questions (Q6, Q7)
- 1 open text question (Q8)
No frequency tables, no summary sheet, no charts, no conditional formatting.
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_010'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

random.seed(42)

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Survey Data ---
    ws = wb.active
    ws.title = "Survey Data"

    # Headers
    headers = [
        "Respondent ID",
        "Q1: Training Content Relevance",
        "Q2: Instructor Effectiveness",
        "Q3: Material Quality",
        "Q4: Practical Applicability",
        "Q5: Overall Satisfaction",
        "Q6: Preferred Training Format",
        "Q7: Would Recommend to Colleague",
        "Q8: Additional Comments",
    ]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 16
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 28
    ws.column_dimensions["G"].width = 28
    ws.column_dimensions["H"].width = 30
    ws.column_dimensions["I"].width = 50

    # Freeze header row
    ws.freeze_panes = "A2"

    # Multiple choice options
    formats = ["Online", "In-Person", "Hybrid"]
    format_weights = [0.35, 0.40, 0.25]
    recommend_options = ["Yes", "No", "Maybe"]
    recommend_weights = [0.55, 0.15, 0.30]

    # Open text comments pool
    comments = [
        "The case studies were very helpful and relevant to our daily work.",
        "Would have preferred more interactive exercises.",
        "Excellent presentation skills by the instructor.",
        "The training could have been shorter - some sections felt repetitive.",
        "I appreciated the real-world examples from recent compliance cases.",
        "The breakout sessions were the highlight for me.",
        "More time should be allocated for Q&A at the end.",
        "The handout materials were comprehensive and well-organized.",
        "Could benefit from more video content to illustrate key points.",
        "The online modules were difficult to navigate on mobile devices.",
        "Great coverage of the new regulatory changes from Q1 2025.",
        "The pre-training assessment was a good way to identify knowledge gaps.",
        "I found the pace too fast for the advanced compliance topics.",
        "The instructor was very knowledgeable and approachable.",
        "Would be better if the training included role-playing scenarios.",
        "The quiz at the end was fair and tested the right concepts.",
        "I suggest adding a follow-up refresher session in 6 months.",
        "The seating arrangement made it hard to see the presentation.",
        "Excellent use of practical examples from our industry.",
        "The training materials should be made available as PDF downloads.",
        "",  # Some respondents leave it blank
        "",
        "",
        "",
        "",
    ]

    # Likert distributions (slightly varied per question to make it realistic)
    # Q1: Training Content Relevance - generally positive
    q1_weights = [0.04, 0.10, 0.20, 0.40, 0.26]
    # Q2: Instructor Effectiveness - very positive
    q2_weights = [0.02, 0.06, 0.16, 0.36, 0.40]
    # Q3: Material Quality - moderate
    q3_weights = [0.06, 0.14, 0.30, 0.32, 0.18]
    # Q4: Practical Applicability - mixed (some find it not applicable)
    q4_weights = [0.10, 0.18, 0.24, 0.30, 0.18]
    # Q5: Overall Satisfaction - generally positive
    q5_weights = [0.04, 0.08, 0.22, 0.38, 0.28]

    def weighted_choice(options, weights):
        return random.choices(options, weights=weights, k=1)[0]

    # Generate 50 respondents
    first_names = [
        "Sarah", "Marcus", "Emily", "James", "Priya", "Michael", "Aisha", "David",
        "Jennifer", "Carlos", "Wei", "Robert", "Maria", "Thomas", "Fatima",
        "Andrew", "Jessica", "Daniel", "Lisa", "Kevin", "Yuki", "Christopher",
        "Amanda", "Brian", "Sophia", "Ryan", "Olivia", "Nathan", "Grace", "Patrick",
        "Hannah", "Steven", "Rachel", "Timothy", "Laura", "Brandon", "Megan",
        "Jonathan", "Ashley", "Gregory", "Nicole", "Aaron", "Stephanie", "Kenneth",
        "Samantha", "Peter", "Victoria", "Douglas", "Rebecca", "Raymond",
    ]
    last_names = [
        "Chen", "Johnson", "Williams", "Garcia", "Patel", "Anderson", "Hassan",
        "Lee", "Martinez", "Thompson", "Zhang", "Davis", "Rodriguez", "Wilson",
        "Ali", "Taylor", "Kim", "Brown", "Jones", "Miller", "Tanaka", "Moore",
        "Clark", "Evans", "Nguyen", "Walker", "Singh", "Hall", "Lopez", "Wright",
        "Adams", "Green", "Baker", "Hill", "Rivera", "Campbell", "Mitchell",
        "Roberts", "Carter", "Phillips", "Turner", "Scott", "Cooper", "Reed",
        "Morgan", "Howard", "Ward", "Torres", "Peterson",
    ]

    for i in range(50):
        row = i + 2
        resp_id = f"RESP-{i+1:03d}"
        ws.cell(row=row, column=1, value=resp_id).alignment = Alignment(horizontal="center")

        # Likert scores (Q1-Q5)
        ws.cell(row=row, column=2, value=weighted_choice([1, 2, 3, 4, 5], q1_weights)).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=weighted_choice([1, 2, 3, 4, 5], q2_weights)).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=weighted_choice([1, 2, 3, 4, 5], q3_weights)).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=weighted_choice([1, 2, 3, 4, 5], q4_weights)).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=6, value=weighted_choice([1, 2, 3, 4, 5], q5_weights)).alignment = Alignment(horizontal="center")

        # Multiple choice
        ws.cell(row=row, column=7, value=weighted_choice(formats, format_weights))
        ws.cell(row=row, column=8, value=weighted_choice(recommend_options, recommend_weights))

        # Open text
        comment = random.choice(comments)
        ws.cell(row=row, column=9, value=comment if comment else None)

        # Light alternating row shading
        if i % 2 == 1:
            light_fill = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
            for c in range(1, 10):
                ws.cell(row=row, column=c).fill = light_fill

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
