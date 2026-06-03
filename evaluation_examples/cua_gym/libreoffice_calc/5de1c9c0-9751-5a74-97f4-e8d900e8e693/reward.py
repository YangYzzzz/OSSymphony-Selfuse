"""
Reward Script: Survey Analysis in LibreOffice Calc
Task ID: calc_grs_010
Domain: libreoffice_calc
Scoring:
  Component 1: Frequency Distribution sheet with COUNTIF formulas (0.25)
  Component 2: Results Summary sheet with AVERAGE formulas (0.20)
  Component 3: Stacked bar chart for Likert questions (0.20)
  Component 4: Pie chart for multiple choice question (0.15)
  Component 5: Conditional formatting on average scores (0.20)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_010'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    sheet_names = wb.sheetnames
    print(f"INFO: Sheets found: {sheet_names}")

    # =========================================================================
    # Component 1: Frequency Distribution sheet with COUNTIF formulas (0.25 pts)
    # This sheet does NOT exist in initial_env — only in golden_env.
    # =========================================================================
    try:
        freq_sheet = None
        for sn in sheet_names:
            if 'frequen' in sn.lower() and 'distri' in sn.lower():
                freq_sheet = wb[sn]
                break

        if freq_sheet is None:
            print("FAIL: Component 1 — No 'Frequency Distribution' sheet found")
        else:
            # Check that we have COUNTIF formulas for 5 Likert questions
            # Each question block has 5 ratings (1-5) with COUNTIF formulas
            countif_count = 0
            percentage_count = 0
            question_blocks = 0

            for row in freq_sheet.iter_rows(min_row=1, max_row=freq_sheet.max_row,
                                            min_col=1, max_col=freq_sheet.max_column,
                                            values_only=False):
                for cell in row:
                    val = cell.value
                    if isinstance(val, str):
                        val_upper = val.upper().replace(" ", "")
                        if 'COUNTIF' in val_upper:
                            countif_count += 1
                        # Count percentage formulas (e.g., =B5/50*100 or similar)
                        if cell.column == 3 and ('/' in val or '%' in val_upper):
                            percentage_count += 1

            # We expect 5 questions x 5 ratings = 25 COUNTIF formulas minimum
            # Also check for question headers
            for row in freq_sheet.iter_rows(min_row=1, max_row=freq_sheet.max_row,
                                            min_col=1, max_col=1, values_only=True):
                val = row[0]
                if isinstance(val, str) and val.startswith('Q') and ':' in val:
                    question_blocks += 1

            print(f"INFO: Component 1 — COUNTIF formulas: {countif_count}, "
                  f"percentage formulas: {percentage_count}, question blocks: {question_blocks}")

            comp1_score = 0.0
            # At least 20 COUNTIF formulas (5 questions x 4+ ratings)
            if countif_count >= 20:
                comp1_score += 0.15
            elif countif_count >= 10:
                comp1_score += 0.08

            # At least 4 question blocks identified
            if question_blocks >= 4:
                comp1_score += 0.10
            elif question_blocks >= 2:
                comp1_score += 0.05

            if comp1_score > 0:
                print(f"PASS: Component 1 — Frequency Distribution with COUNTIF "
                      f"({comp1_score} pts)")
                total_score += comp1_score
            else:
                print("FAIL: Component 1 — Insufficient COUNTIF formulas or question blocks")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Results Summary sheet with AVERAGE formulas (0.20 pts)
    # This sheet does NOT exist in initial_env — only in golden_env.
    # =========================================================================
    try:
        summary_sheet = None
        for sn in sheet_names:
            sn_lower = sn.lower()
            if 'summary' in sn_lower or 'result' in sn_lower:
                summary_sheet = wb[sn]
                break

        if summary_sheet is None:
            print("FAIL: Component 2 — No 'Results Summary' sheet found")
        else:
            average_count = 0
            question_labels = 0

            for row in summary_sheet.iter_rows(min_row=1, max_row=summary_sheet.max_row,
                                               min_col=1, max_col=summary_sheet.max_column,
                                               values_only=False):
                for cell in row:
                    val = cell.value
                    if isinstance(val, str):
                        val_upper = val.upper().replace(" ", "")
                        if 'AVERAGE' in val_upper:
                            average_count += 1
                        # Count Likert question labels in column A
                        if cell.column == 1 and val.startswith('Q') and ':' in val:
                            question_labels += 1

            print(f"INFO: Component 2 — AVERAGE formulas: {average_count}, "
                  f"question labels: {question_labels}")

            comp2_score = 0.0
            # At least 5 AVERAGE formulas (one per Likert question)
            if average_count >= 5:
                comp2_score += 0.12
            elif average_count >= 3:
                comp2_score += 0.06

            # At least 4 question labels
            if question_labels >= 4:
                comp2_score += 0.08
            elif question_labels >= 2:
                comp2_score += 0.04

            if comp2_score > 0:
                print(f"PASS: Component 2 — Results Summary with AVERAGE "
                      f"({comp2_score} pts)")
                total_score += comp2_score
            else:
                print("FAIL: Component 2 — Insufficient AVERAGE formulas or question labels")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Stacked bar chart for Likert question distribution (0.20 pts)
    # No charts exist in initial_env.
    # =========================================================================
    try:
        bar_chart_found = False
        bar_chart_stacked = False
        bar_chart_series_ok = False

        # Search all sheets for charts
        for sn in sheet_names:
            ws = wb[sn]
            for ch in ws._charts:
                class_name = ch.__class__.__name__
                if class_name in ('BarChart', 'BarChart3D'):
                    bar_chart_found = True
                    # Check if stacked
                    grouping = getattr(ch, 'grouping', None)
                    if grouping == 'stacked' or grouping == 'percentStacked':
                        bar_chart_stacked = True
                    # Check series count (should have ~5 for 5 rating levels)
                    if len(ch.series) >= 3:
                        bar_chart_series_ok = True
                    print(f"INFO: Component 3 — Bar chart found: type={getattr(ch, 'type', 'N/A')}, "
                          f"grouping={grouping}, series={len(ch.series)}")

        comp3_score = 0.0
        if bar_chart_found:
            comp3_score += 0.08
        if bar_chart_stacked:
            comp3_score += 0.06
        if bar_chart_series_ok:
            comp3_score += 0.06

        if comp3_score > 0:
            print(f"PASS: Component 3 — Bar chart ({comp3_score} pts)")
            total_score += comp3_score
        else:
            print("FAIL: Component 3 — No bar chart found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: Pie chart for a multiple choice question (0.15 pts)
    # No charts exist in initial_env.
    # =========================================================================
    try:
        pie_chart_found = False
        pie_chart_series_ok = False

        for sn in sheet_names:
            ws = wb[sn]
            for ch in ws._charts:
                class_name = ch.__class__.__name__
                if class_name in ('PieChart', 'PieChart3D'):
                    pie_chart_found = True
                    if len(ch.series) >= 1:
                        pie_chart_series_ok = True
                    print(f"INFO: Component 4 — Pie chart found, series={len(ch.series)}")

        comp4_score = 0.0
        if pie_chart_found:
            comp4_score += 0.10
        if pie_chart_series_ok:
            comp4_score += 0.05

        if comp4_score > 0:
            print(f"PASS: Component 4 — Pie chart ({comp4_score} pts)")
            total_score += comp4_score
        else:
            print("FAIL: Component 4 — No pie chart found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # =========================================================================
    # Component 5: Conditional formatting on average scores (0.20 pts)
    # Task: highlight avg < 3.0 in red, avg > 4.0 in green
    # No conditional formatting exists in initial_env.
    # =========================================================================
    try:
        cf_found = False
        has_red_rule = False
        has_green_rule = False

        # Search the Results Summary sheet for conditional formatting
        if summary_sheet is not None:
            for cf in summary_sheet.conditional_formatting:
                cf_found = True
                for rule in cf.rules:
                    rule_type = getattr(rule, 'type', '')
                    rule_operator = getattr(rule, 'operator', '')
                    rule_formula = getattr(rule, 'formula', [])

                    # Check for red rule (less than 3.0)
                    if rule_operator in ('lessThan', 'lessThanOrEqual'):
                        if rule.dxf and rule.dxf.fill and rule.dxf.fill.fgColor:
                            color = rule.dxf.fill.fgColor.rgb
                            if color and 'FF' in color[:4]:
                                # Red-ish fill (contains FF in red channel)
                                has_red_rule = True
                                print(f"INFO: Component 5 — Red rule: op={rule_operator}, "
                                      f"formula={rule_formula}, fill={color}")

                    # Check for green rule (greater than 4.0)
                    if rule_operator in ('greaterThan', 'greaterThanOrEqual'):
                        if rule.dxf and rule.dxf.fill and rule.dxf.fill.fgColor:
                            color = rule.dxf.fill.fgColor.rgb
                            if color:
                                has_green_rule = True
                                print(f"INFO: Component 5 — Green rule: op={rule_operator}, "
                                      f"formula={rule_formula}, fill={color}")

        comp5_score = 0.0
        if cf_found:
            comp5_score += 0.05
        if has_red_rule:
            comp5_score += 0.075
        if has_green_rule:
            comp5_score += 0.075

        if comp5_score > 0:
            print(f"PASS: Component 5 — Conditional formatting ({comp5_score} pts)")
            total_score += comp5_score
        else:
            print("FAIL: Component 5 — No conditional formatting found on average scores")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
