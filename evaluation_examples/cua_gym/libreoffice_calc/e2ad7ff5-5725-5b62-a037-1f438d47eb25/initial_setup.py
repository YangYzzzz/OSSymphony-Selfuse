"""
Initial Setup: Product Price List (unformatted)
Task ID: calc_sales_pricing_list_format_037
Domain: libreoffice_calc

Creates PriceList sheet with 50 product rows.
No merged cells, no borders, no alternating shading,
no currency format, no print area set — all to be added by agent.
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_pricing_list_format_037'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'PriceList'

    # Row 1: title (NOT merged, NOT formatted)
    ws['A1'] = 'Product Price List 2024'

    # Row 2: headers (NOT formatted)
    headers = ['SKU', 'Product Name', 'Category', 'List Price', 'Min Order', 'Lead Time']
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)

    # Rows 3-52: 50 realistic products (prices as plain numbers, no currency format)
    products = [
        ('SKU-001', 'Ergonomic Office Chair',       'Furniture',      549.99,  1, '5-7 days'),
        ('SKU-002', 'Standing Desk 60"',             'Furniture',     1249.00,  1, '7-10 days'),
        ('SKU-003', 'Monitor Arm Dual',              'Accessories',    189.50,  2, '3-5 days'),
        ('SKU-004', 'Mechanical Keyboard TKL',       'Electronics',    129.95,  5, '2-4 days'),
        ('SKU-005', 'Wireless Mouse Pro',            'Electronics',     79.99, 10, '2-4 days'),
        ('SKU-006', 'USB-C Docking Station',         'Electronics',    249.00,  2, '3-5 days'),
        ('SKU-007', 'LED Desk Lamp 24W',             'Lighting',        59.95,  5, '3-5 days'),
        ('SKU-008', 'Cable Management Tray',         'Accessories',     29.99, 10, '2-3 days'),
        ('SKU-009', 'Laptop Stand Aluminium',        'Accessories',     89.50,  5, '2-4 days'),
        ('SKU-010', 'Monitor 27" 4K IPS',            'Electronics',    699.00,  1, '5-7 days'),
        ('SKU-011', 'Noise-Cancelling Headset',      'Electronics',    349.99,  2, '3-5 days'),
        ('SKU-012', 'Webcam 1080p HD',               'Electronics',    119.95,  5, '2-4 days'),
        ('SKU-013', 'Desk Organiser Set',            'Accessories',     39.99, 10, '2-3 days'),
        ('SKU-014', 'Whiteboard 48x36"',             'Office Supplies', 149.00,  2, '5-7 days'),
        ('SKU-015', 'Printer Laser Mono',            'Electronics',    429.00,  1, '5-7 days'),
        ('SKU-016', 'Paper Ream A4 500-sht',         'Office Supplies',  12.99, 50, '1-2 days'),
        ('SKU-017', 'Stapler Heavy Duty',            'Office Supplies',  24.95, 10, '1-2 days'),
        ('SKU-018', 'Filing Cabinet 4-Draw',         'Furniture',      389.00,  1, '7-10 days'),
        ('SKU-019', 'Bookcase 5-Shelf',              'Furniture',      279.99,  1, '7-10 days'),
        ('SKU-020', 'Visitor Chair Padded',          'Furniture',      199.00,  4, '5-7 days'),
        ('SKU-021', 'Conference Table 8-seat',       'Furniture',     1899.00,  1, '14-21 days'),
        ('SKU-022', 'Projector 3500 Lumen',          'Electronics',    899.00,  1, '5-7 days'),
        ('SKU-023', 'Projection Screen 100"',        'Electronics',    329.00,  1, '5-7 days'),
        ('SKU-024', 'Wi-Fi Router Tri-band',         'Electronics',    249.95,  2, '3-5 days'),
        ('SKU-025', 'Network Switch 24-Port',        'Electronics',    399.00,  1, '3-5 days'),
        ('SKU-026', 'UPS 1500VA',                    'Electronics',    279.99,  1, '3-5 days'),
        ('SKU-027', 'External SSD 2TB',              'Electronics',    189.95,  5, '2-4 days'),
        ('SKU-028', 'Flash Drive 128GB Pack',        'Electronics',     49.99, 10, '1-2 days'),
        ('SKU-029', 'Shredder Cross-Cut P-4',        'Electronics',    169.00,  2, '3-5 days'),
        ('SKU-030', 'Paper Folder Electric',         'Electronics',    299.00,  1, '5-7 days'),
        ('SKU-031', 'Toner Cartridge Black',         'Consumables',     64.99, 10, '1-2 days'),
        ('SKU-032', 'Ink Cartridge Colour Set',      'Consumables',     39.95, 10, '1-2 days'),
        ('SKU-033', 'Binding Machine Comb',          'Office Supplies', 129.00,  2, '3-5 days'),
        ('SKU-034', 'Laminator A3',                  'Office Supplies', 149.95,  2, '3-5 days'),
        ('SKU-035', 'Label Printer Thermo',          'Electronics',     99.00,  5, '2-4 days'),
        ('SKU-036', 'Label Roll 62mm 10-pk',         'Consumables',     34.99, 20, '1-2 days'),
        ('SKU-037', 'Ergonomic Footrest',            'Accessories',     69.95,  5, '3-5 days'),
        ('SKU-038', 'Monitor Privacy Filter 27"',    'Accessories',     79.00,  5, '2-4 days'),
        ('SKU-039', 'Bluetooth Speaker Compact',     'Electronics',     89.99, 10, '2-4 days'),
        ('SKU-040', 'Smart Plug Power Strip',        'Electronics',     59.95, 10, '2-3 days'),
        ('SKU-041', 'Desk Pad Large 80x40cm',        'Accessories',     44.99, 10, '2-3 days'),
        ('SKU-042', 'Pen Holder Mesh Steel',         'Accessories',     17.99, 20, '1-2 days'),
        ('SKU-043', 'Notebook Spiral A4 pk5',        'Office Supplies',  22.50, 20, '1-2 days'),
        ('SKU-044', 'Ball Pen Blue Box-50',          'Office Supplies',  14.99, 50, '1-2 days'),
        ('SKU-045', 'Highlighter Set 5col',          'Office Supplies',   9.99, 20, '1-2 days'),
        ('SKU-046', 'Scissors Stainless 8"',         'Office Supplies',  12.50, 20, '1-2 days'),
        ('SKU-047', 'Tape Dispenser + 3 Rolls',      'Office Supplies',  15.95, 20, '1-2 days'),
        ('SKU-048', 'Sticky Notes 3x3 pk12',         'Office Supplies',  18.99, 20, '1-2 days'),
        ('SKU-049', 'Binder A4 50mm Blue pk5',       'Office Supplies',  29.95, 10, '1-2 days'),
        ('SKU-050', 'Document Wallet Zip A4 pk10',   'Office Supplies',  24.99, 20, '1-2 days'),
    ]

    for row_idx, (sku, name, category, price, min_ord, lead) in enumerate(products, 3):
        ws.cell(row=row_idx, column=1, value=sku)
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=category)
        ws.cell(row=row_idx, column=4, value=price)   # plain number, no currency format
        ws.cell(row=row_idx, column=5, value=min_ord)
        ws.cell(row=row_idx, column=6, value=lead)

    # Set reasonable column widths for readability but NO formatting
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
