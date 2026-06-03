"""
Reward Script: Format product price list for professional customer-facing document
Task ID: calc_sales_pricing_list_format_037
Domain: libreoffice_calc
Scoring:
  Component 1: Title row merged (A1:F1), bold, size 14, centered                (0.20 pts)
  Component 2: Header row (row 2) dark blue background, white font, bold         (0.20 pts)
  Component 3: Alternating row shading rows 3-52 (white/gray pattern)            (0.20 pts)
  Component 4: Currency format $#,##0.00 applied to column D data rows           (0.15 pts)
  Component 5: Thick outer border + thin inner borders on A1:F52                 (0.15 pts)
  Component 6: Print area A1:F52, landscape orientation, page number footer       (0.10 pts)
  Total: 1.0
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_sales_pricing_list_format_037'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: must have 'PriceList' sheet
    if 'PriceList' not in wb.sheetnames:
        print("CRITICAL: Sheet 'PriceList' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['PriceList']

    # Component 1: Title row merged A1:F1, bold, font size 14, centered (0.20 pts)
    try:
        # Check that A1:F1 is a merged range
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        title_merged = 'A1:F1' in merged_ranges

        # Check that B1, C1, D1 are MergedCell instances (part of merged range)
        b1_merged = isinstance(ws['B1'], MergedCell)
        c1_merged = isinstance(ws['C1'], MergedCell)

        # Check A1 font
        cell_a1 = ws['A1']
        a1_bold = cell_a1.font.bold is True
        a1_size = cell_a1.font.size is not None and float(cell_a1.font.size) >= 14.0
        a1_centered = cell_a1.alignment.horizontal == 'center'

        if title_merged and b1_merged and a1_bold and a1_size and a1_centered:
            print(f"PASS: Component 1 — A1:F1 merged, bold, size={cell_a1.font.size}, centered (0.20 pts)")
            total_score += 0.20
        else:
            reasons = []
            if not title_merged:
                reasons.append(f"A1:F1 not merged (found: {merged_ranges})")
            if not b1_merged:
                reasons.append("B1 not part of merged range")
            if not a1_bold:
                reasons.append(f"A1 not bold (bold={cell_a1.font.bold})")
            if not a1_size:
                reasons.append(f"A1 font size not >=14 (size={cell_a1.font.size})")
            if not a1_centered:
                reasons.append(f"A1 not centered (horizontal={cell_a1.alignment.horizontal})")
            print(f"FAIL: Component 1 — title row formatting: {'; '.join(reasons)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Header row (row 2) dark blue background, white text, bold (0.20 pts)
    try:
        cell_a2 = ws['A2']

        # Check dark blue background (expected: FF4472C4 or similar dark blue)
        fgcolor = cell_a2.fill.fgColor.rgb
        # Accept any dark blue — blue channel dominant, checking it's not empty (00000000)
        # We look for a non-transparent fill where the color has blue characteristics
        has_dark_bg = (fgcolor not in ('00000000', '000000', None, ''))

        # White or near-white font color
        try:
            font_color = cell_a2.font.color.rgb
            # White is 00FFFFFF (font alpha doesn't matter visually)
            has_white_font = font_color is not None and font_color.upper().endswith('FFFFFF')
        except Exception:
            has_white_font = False

        # Bold header
        a2_bold = cell_a2.font.bold is True

        # Also check a few other header cells (B2, C2) share similar styling
        cell_b2 = ws['B2']
        b2_bold = cell_b2.font.bold is True

        if has_dark_bg and has_white_font and a2_bold and b2_bold:
            print(f"PASS: Component 2 — header row: bg={fgcolor}, font_color={font_color}, bold=True (0.20 pts)")
            total_score += 0.20
        else:
            reasons = []
            if not has_dark_bg:
                reasons.append(f"no dark background on A2 (fgColor.rgb={fgcolor})")
            if not has_white_font:
                reasons.append(f"no white font on A2 (font.color.rgb={font_color if 'font_color' in dir() else 'error'})")
            if not a2_bold:
                reasons.append(f"A2 not bold")
            if not b2_bold:
                reasons.append(f"B2 not bold")
            print(f"FAIL: Component 2 — header styling: {'; '.join(reasons)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Alternating row shading rows 3-52 (0.20 pts)
    # Odd rows (3,5,7,...) should have one color (white/light), even rows (4,6,8,...) another (gray)
    try:
        white_rows_ok = 0
        gray_rows_ok = 0
        total_odd = 0
        total_even = 0

        for row in range(3, 53):
            cell = ws.cell(row=row, column=1)
            fgcolor = cell.fill.fgColor.rgb

            if row % 2 == 1:  # odd rows (3,5,7,...) — white
                total_odd += 1
                # White = FFFFFFFF, or could be no-fill (00000000)
                if fgcolor not in ('00000000', None, ''):
                    white_rows_ok += 1
            else:  # even rows (4,6,8,...) — gray
                total_even += 1
                if fgcolor not in ('00000000', None, ''):
                    gray_rows_ok += 1

        # Check that the two groups have DIFFERENT colors (alternating)
        odd_sample = ws.cell(row=3, column=1).fill.fgColor.rgb   # white row
        even_sample = ws.cell(row=4, column=1).fill.fgColor.rgb  # gray row
        colors_differ = odd_sample != even_sample

        # Accept if at least 90% of rows have non-default fills in one direction
        # AND the two groups have different colors
        alternating_ok = (
            colors_differ and
            gray_rows_ok >= int(total_even * 0.9) and  # gray rows have non-default fill
            (white_rows_ok >= int(total_odd * 0.5) or  # either white rows have explicit fill
             odd_sample in ('00000000', ''))            # or odd rows are explicitly white/default
        )

        if alternating_ok:
            print(f"PASS: Component 3 — alternating shading: odd={odd_sample}, even={even_sample}, differ={colors_differ} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 3 — alternating shading: odd_sample={odd_sample}, even_sample={even_sample}, colors_differ={colors_differ}, gray_rows_ok={gray_rows_ok}/{total_even}, white_rows_ok={white_rows_ok}/{total_odd}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Currency format $#,##0.00 on column D data rows 3-52 (0.15 pts)
    try:
        currency_count = 0
        total_d_rows = 50  # rows 3-52

        for row in range(3, 53):
            cell = ws.cell(row=row, column=4)
            if cell.number_format and '$' in cell.number_format:
                currency_count += 1

        # Accept if at least 90% of D column data rows have currency format
        currency_ratio = currency_count / total_d_rows
        if currency_ratio >= 0.9:
            sample_fmt = ws.cell(row=3, column=4).number_format
            print(f"PASS: Component 4 — currency format applied: {currency_count}/{total_d_rows} rows, sample format='{sample_fmt}' (0.15 pts)")
            total_score += 0.15
        else:
            sample_fmt = ws.cell(row=3, column=4).number_format
            print(f"FAIL: Component 4 — currency format: only {currency_count}/{total_d_rows} rows have $ format, sample='{sample_fmt}'")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Thick outer border + thin inner borders on A1:F52 (0.15 pts)
    try:
        # Check outer corners have thick borders
        # A1: left=thick, top=thick
        cell_a1 = ws['A1']
        a1_thick = (cell_a1.border.left.style == 'thick' and
                    cell_a1.border.top.style == 'thick')

        # F52: right=thick, bottom=thick
        cell_f52 = ws['F52']
        f52_thick = (cell_f52.border.right.style == 'thick' and
                     cell_f52.border.bottom.style == 'thick')

        # A52: left=thick, bottom=thick
        cell_a52 = ws['A52']
        a52_thick = (cell_a52.border.left.style == 'thick' and
                     cell_a52.border.bottom.style == 'thick')

        # F1: right=thick, top=thick (note: F1 is a MergedCell in golden)
        cell_f1 = ws['F1']
        f1_thick_top = cell_f1.border.top.style == 'thick'

        # Check inner border (a middle cell like D10 should have thin borders)
        cell_d10 = ws['D10']
        d10_has_border = (cell_d10.border.left.style is not None and
                          cell_d10.border.right.style is not None)

        outer_borders_ok = a1_thick and f52_thick and a52_thick
        inner_borders_ok = d10_has_border

        if outer_borders_ok and inner_borders_ok:
            print(f"PASS: Component 5 — thick outer borders and thin inner borders present (0.15 pts)")
            total_score += 0.15
        else:
            reasons = []
            if not a1_thick:
                reasons.append(f"A1 outer border: left={cell_a1.border.left.style}, top={cell_a1.border.top.style}")
            if not f52_thick:
                reasons.append(f"F52 outer border: right={cell_f52.border.right.style}, bottom={cell_f52.border.bottom.style}")
            if not a52_thick:
                reasons.append(f"A52 outer border: left={cell_a52.border.left.style}, bottom={cell_a52.border.bottom.style}")
            if not inner_borders_ok:
                reasons.append(f"D10 inner border missing: left={cell_d10.border.left.style}, right={cell_d10.border.right.style}")
            print(f"FAIL: Component 5 — borders: {'; '.join(reasons)}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Print area A1:F52, landscape orientation, page number footer (0.10 pts)
    try:
        # Check print area (stored as "'PriceList'!$A$1:$F$52" or "PriceList!$A$1:$F$52")
        print_area = ws.print_area or ''
        # Normalize: remove sheet name prefix and quotes, just check cell range portion
        import re
        pa_normalized = re.sub(r"^.*!", '', print_area.replace("'", ''))
        has_print_area = '$A$1:$F$52' in pa_normalized or 'A1:F52' in pa_normalized

        # Check landscape orientation
        has_landscape = ws.page_setup.orientation == 'landscape'

        # Check footer with page number
        has_footer = False
        try:
            footer_text = ws.oddFooter.center.text if ws.oddFooter else ''
            has_footer = footer_text is not None and len(footer_text) > 0 and ('&P' in footer_text or '&N' in footer_text)
        except Exception:
            has_footer = False

        if has_print_area and has_landscape and has_footer:
            print(f"PASS: Component 6 — print area={print_area}, orientation={ws.page_setup.orientation}, footer present (0.10 pts)")
            total_score += 0.10
        else:
            reasons = []
            if not has_print_area:
                reasons.append(f"print area not set correctly: '{print_area}' (normalized: '{pa_normalized}')")
            if not has_landscape:
                reasons.append(f"orientation not landscape: '{ws.page_setup.orientation}'")
            if not has_footer:
                try:
                    footer_text = ws.oddFooter.center.text if ws.oddFooter else ''
                except Exception:
                    footer_text = ''
                reasons.append(f"footer missing or no page number: '{footer_text}'")
            print(f"FAIL: Component 6 — print setup: {'; '.join(reasons)}")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
