"""
Initial Setup: Hide formulas, unlock student cells, protect sheet
Task ID: calc_ps_032
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_032'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Exam'

    # --- Row 1: Weighted scoring formulas (A1:J1) ---
    # These are the formulas the instructor wants to hide
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Column headers with weighted score formulas
    # Each formula computes a weighted score for a category
    formula_headers = [
        ("A1", "=0.15*SUM(A2:A50)", "Multiple Choice (15%)"),
        ("B1", "=0.10*SUM(B2:B50)", "True/False (10%)"),
        ("C1", "=0.20*SUM(C2:C50)", "Short Answer (20%)"),
        ("D1", "=0.15*SUM(D2:D50)", "Fill in Blank (15%)"),
        ("E1", "=0.10*SUM(E2:E50)", "Matching (10%)"),
        ("F1", "=0.05*SUM(F2:F50)", "Bonus (5%)"),
        ("G1", "=0.10*SUM(G2:G50)", "Lab Practical (10%)"),
        ("H1", "=0.05*SUM(H2:H50)", "Attendance (5%)"),
        ("I1", "=0.10*SUM(I2:I50)", "Essay (10%)"),
        ("J1", "=SUM(A1:I1)", "Total Weighted Score"),
    ]

    for coord, formula, _ in formula_headers:
        cell = ws[coord]
        cell.value = formula
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        # All cells locked by default (Protection(locked=True)), NOT hidden
        cell.protection = Protection(locked=True, hidden=False)

    # Set column widths
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws.column_dimensions[col_letter].width = 18

    # --- Rows 2-11: Sample student response data ---
    students_data = [
        [8, 9, 15, 12, 8, 3, 7, 4, 8, None],
        [7, 8, 18, 10, 9, 5, 9, 5, 7, None],
        [9, 10, 12, 14, 7, 2, 8, 3, 9, None],
        [6, 7, 16, 11, 10, 4, 6, 5, 6, None],
        [10, 9, 19, 13, 8, 5, 10, 4, 10, None],
        [5, 6, 11, 9, 6, 1, 5, 3, 5, None],
        [8, 10, 17, 12, 9, 3, 8, 5, 8, None],
        [7, 8, 14, 10, 7, 4, 7, 4, 7, None],
        [9, 9, 20, 15, 10, 5, 9, 5, 9, None],
        [6, 7, 13, 11, 8, 2, 6, 3, 6, None],
    ]

    for r, row_data in enumerate(students_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c)
            if val is not None:
                cell.value = val
            # All cells locked by default, NOT hidden
            cell.protection = Protection(locked=True, hidden=False)

    # Rows 12-50: empty but with default protection (locked, not hidden)
    for r in range(12, 51):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.protection = Protection(locked=True, hidden=False)

    # Sheet is NOT protected in initial state
    # (ws.protection.sheet is False by default)

    ws.row_dimensions[1].height = 30

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
