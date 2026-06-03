"""
Reward Script: Hide formulas in row 1, unlock A2:J50, protect sheet with password
Task ID: calc_ps_032
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.35): A1:J1 cells have hidden=True (formula hiding)
  - Component 2 (0.35): A2:J50 cells are unlocked (locked=False)
  - Component 3 (0.30): Sheet is protected with password 'exam2024'
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_032'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify 'Exam' sheet exists
    if 'Exam' not in wb.sheetnames:
        print("CRITICAL: 'Exam' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Exam']

    # Component 1: A1:J1 cells have hidden=True (0.35 points)
    # In the initial state, these are hidden=False. The task requires hiding formulas.
    try:
        hidden_count = 0
        total_row1 = 10  # columns A through J
        for col in range(1, 11):
            cell = ws.cell(row=1, column=col)
            if cell.protection.hidden:
                hidden_count += 1

        if hidden_count == total_row1:
            print(f"PASS: Component 1 — All {total_row1} cells in A1:J1 have hidden=True (0.35 pts)")
            total_score += 0.35
        elif hidden_count > 0:
            partial = 0.35 * (hidden_count / total_row1)
            print(f"PARTIAL: Component 1 — {hidden_count}/{total_row1} cells in row 1 hidden ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No cells in A1:J1 have hidden=True (0/{total_row1})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: A2:J50 cells are unlocked (0.35 points)
    # In the initial state, all cells are locked=True. The task requires unlocking A2:J50.
    try:
        unlocked_count = 0
        total_cells = 0
        for row in range(2, 51):
            for col in range(1, 11):
                total_cells += 1
                cell = ws.cell(row=row, column=col)
                if not cell.protection.locked:
                    unlocked_count += 1

        if unlocked_count == total_cells:
            print(f"PASS: Component 2 — All {total_cells} cells in A2:J50 are unlocked (0.35 pts)")
            total_score += 0.35
        elif unlocked_count > 0:
            partial = 0.35 * (unlocked_count / total_cells)
            print(f"PARTIAL: Component 2 — {unlocked_count}/{total_cells} cells unlocked ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No cells in A2:J50 are unlocked (0/{total_cells})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Sheet is protected with password 'exam2024' (0.30 points)
    # In the initial state, sheet is not protected. The task requires protection with specific password.
    try:
        if ws.protection.sheet:
            # Sheet is protected - check password hash
            # openpyxl hashes 'exam2024' to 'EDF1'
            expected_hash = openpyxl.worksheet.protection.hash_password('exam2024')
            actual_hash = ws.protection.password

            if actual_hash == expected_hash:
                print(f"PASS: Component 3 — Sheet protected with correct password hash (0.30 pts)")
                total_score += 0.30
            elif actual_hash is not None:
                print(f"PARTIAL: Component 3 — Sheet protected but password hash mismatch "
                      f"(expected={expected_hash}, actual={actual_hash}) (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 3 — Sheet protected but no password set")
        else:
            print(f"FAIL: Component 3 — Sheet is not protected")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
