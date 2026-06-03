"""
Reward Script: Create book_club.xlsx spreadsheet from Goodreads data
Task ID: osworld_multi_apps_misc_003
Domain: libreoffice_calc (multi-app: Chrome + LibreOffice Writer + LibreOffice Calc)
Scoring:
  Component 1: Correct headers (Title, Author, Average Rating, Number of Ratings) — 0.3 pts
  Component 2: All 5 books present with correct titles — 0.3 pts
  Component 3: All 5 books have correct authors — 0.2 pts
  Component 4: Average Rating and Number of Ratings populated as numeric values — 0.2 pts
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_misc_003'
FILE_PATH = '/home/user/Desktop/book_club.xlsx'

# Expected book titles (case-insensitive comparison used)
EXPECTED_TITLES = {
    '1984',
    'brave new world',
    'fahrenheit 451',
    "the handmaid's tale",
    'never let me go',
}

# Expected authors mapped to normalized title (lowercase)
EXPECTED_AUTHORS = {
    '1984': 'george orwell',
    'brave new world': 'aldous huxley',
    'fahrenheit 451': 'ray bradbury',
    "the handmaid's tale": 'margaret atwood',
    'never let me go': 'kazuo ishiguro',
}

# Expected headers (case-insensitive comparison)
EXPECTED_HEADERS = ['title', 'author', 'average rating', 'number of ratings']


def normalize(s):
    """Normalize a string for comparison."""
    if s is None:
        return ''
    return str(s).strip().lower()


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Gate: file must exist to proceed
    if not os.path.exists(file_path):
        print(f"CRITICAL: File not found at {file_path}")
        print("REWARD: 0.0")
        return 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load workbook {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get first sheet (task says 'a new spreadsheet', don't restrict sheet name)
    try:
        ws = wb.worksheets[0]
        print(f"Loaded sheet: {ws.title}, rows: {ws.max_row}, cols: {ws.max_column}")
    except Exception as e:
        print(f"CRITICAL: Cannot access worksheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # ---------------------------------------------------------------------------
    # Component 1: Correct headers (Title, Author, Average Rating, Number of Ratings)
    # 0.3 points — verifies that the file was created with the correct column structure
    # ---------------------------------------------------------------------------
    try:
        actual_headers = []
        for c in range(1, 5):
            val = ws.cell(row=1, column=c).value
            actual_headers.append(normalize(val))

        headers_ok = (actual_headers == EXPECTED_HEADERS)

        if headers_ok:
            print(f"PASS: Component 1 — headers correct: {actual_headers} (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — expected headers {EXPECTED_HEADERS}, found: {actual_headers}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ---------------------------------------------------------------------------
    # Component 2: All 5 books present with correct titles
    # 0.3 points — verifies that all 5 books from the reading list were included
    # ---------------------------------------------------------------------------
    try:
        found_titles = set()
        for row_idx in range(2, ws.max_row + 1):
            title_val = ws.cell(row=row_idx, column=1).value
            if title_val is not None:
                found_titles.add(normalize(title_val))

        matched_titles = EXPECTED_TITLES.intersection(found_titles)
        num_matched = len(matched_titles)

        if num_matched == 5:
            print(f"PASS: Component 2 — all 5 book titles found (0.3 pts)")
            total_score += 0.3
        elif num_matched >= 3:
            partial = 0.15
            print(f"PARTIAL: Component 2 — {num_matched}/5 titles found, partial credit (0.15 pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — only {num_matched}/5 titles found. Found: {found_titles}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---------------------------------------------------------------------------
    # Component 3: All 5 books have correct authors
    # 0.2 points — verifies that author data was correctly populated from Goodreads
    # ---------------------------------------------------------------------------
    try:
        author_matches = 0
        for row_idx in range(2, ws.max_row + 1):
            title_val = ws.cell(row=row_idx, column=1).value
            author_val = ws.cell(row=row_idx, column=2).value

            title_norm = normalize(title_val)
            author_norm = normalize(author_val)

            if title_norm in EXPECTED_AUTHORS:
                expected_author = EXPECTED_AUTHORS[title_norm]
                if expected_author in author_norm or author_norm in expected_author:
                    author_matches += 1

        if author_matches == 5:
            print(f"PASS: Component 3 — all 5 authors correct (0.2 pts)")
            total_score += 0.2
        elif author_matches >= 3:
            partial = 0.1
            print(f"PARTIAL: Component 3 — {author_matches}/5 authors correct, partial credit (0.1 pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — only {author_matches}/5 authors correct")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ---------------------------------------------------------------------------
    # Component 4: Average Rating and Number of Ratings populated as numeric values
    # 0.2 points — verifies that numeric data from Goodreads was captured
    # ---------------------------------------------------------------------------
    try:
        numeric_rows = 0
        for row_idx in range(2, ws.max_row + 1):
            title_val = ws.cell(row=row_idx, column=1).value
            if title_val is None:
                continue

            avg_rating = ws.cell(row=row_idx, column=3).value
            num_ratings = ws.cell(row=row_idx, column=4).value

            # Check that both values are numeric and within realistic ranges
            avg_valid = False
            num_valid = False

            try:
                avg_float = float(avg_rating)
                # Goodreads ratings are between 1.0 and 5.0
                avg_valid = (1.0 <= avg_float <= 5.0)
            except (TypeError, ValueError):
                avg_valid = False

            try:
                num_int = int(float(num_ratings))
                # A real book should have at least 1 rating on Goodreads
                num_valid = (num_int > 0)
            except (TypeError, ValueError):
                num_valid = False

            if avg_valid and num_valid:
                numeric_rows += 1

        if numeric_rows == 5:
            print(f"PASS: Component 4 — all 5 rows have valid numeric rating data (0.2 pts)")
            total_score += 0.2
        elif numeric_rows >= 3:
            partial = 0.1
            print(f"PARTIAL: Component 4 — {numeric_rows}/5 rows have valid numeric rating data, partial credit (0.1 pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — only {numeric_rows}/5 rows have valid numeric rating data")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point: verify the task artifact at canonical path
if not os.path.exists(FILE_PATH):
    print(f"File not found: {FILE_PATH}")
    print("REWARD: 0.0")
else:
    verify_task(FILE_PATH)
