"""
Reward Script: AutoFill Month Name Series in LibreOffice Calc
Task ID: calc_cop_autofill_004
Domain: libreoffice_calc

Task: Use AutoFill to extend the month name series starting with 'January' in A2
and 'February' in A3 down to A13 to complete all 12 months.

Scoring:
  Component 1: A4:A8 filled with correct month names (March through July) — 0.5 pts
  Component 2: A9:A13 filled with correct month names (August through December) — 0.5 pts
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_cop_autofill_004'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    The task requires AutoFill of months in A4:A13 (10 new cells).
    Initial file has only A2='January', A3='February'.
    Golden file has all 12 months in A2:A13.
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the sheet 'MonthlyReport' exists (precondition gate)
    if 'MonthlyReport' not in wb.sheetnames:
        print("CRITICAL: Sheet 'MonthlyReport' not found.")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['MonthlyReport']

    # Component 1: A4:A8 contain correct month names March through July (0.5 points)
    # These cells are EMPTY in the initial file, so any correct value means task progress.
    # Checks that the autofill produced the right first-half sequence.
    try:
        months_first_half = ['March', 'April', 'May', 'June', 'July']
        mismatches_first = []

        for i, expected_month in enumerate(months_first_half):
            row = i + 4  # A4 to A8
            actual = ws.cell(row=row, column=1).value
            if actual != expected_month:
                mismatches_first.append(f"A{row}: expected '{expected_month}', found {repr(actual)}")

        if len(mismatches_first) == 0:
            print("PASS: Component 1 — A4:A8 contain March through July (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — A4:A8 months incorrect: {', '.join(mismatches_first)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: A9:A13 contain correct month names August through December (0.5 points)
    # These cells are also EMPTY in the initial file.
    # Checks that the autofill produced the right second-half sequence.
    try:
        months_second_half = ['August', 'September', 'October', 'November', 'December']
        mismatches_second = []

        for i, expected_month in enumerate(months_second_half):
            row = i + 9  # A9 to A13
            actual = ws.cell(row=row, column=1).value
            if actual != expected_month:
                mismatches_second.append(f"A{row}: expected '{expected_month}', found {repr(actual)}")

        if len(mismatches_second) == 0:
            print("PASS: Component 2 — A9:A13 contain August through December (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 2 — A9:A13 months incorrect: {', '.join(mismatches_second)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
