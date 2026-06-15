"""
Initial Setup: AutoFill month name series task
Task ID: calc_cop_autofill_004
Domain: libreoffice_calc

Creates a MonthlyReport spreadsheet where:
- A2: 'January', A3: 'February', A4:A13: empty (awaiting AutoFill)
- Columns B-D contain realistic monthly sales data for the first two months
- The agent must use AutoFill to extend months March-December in A4:A13
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_cop_autofill_004'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'MonthlyReport'

    # --- Header row ---
    headers = ['Month', 'Revenue ($)', 'Expenses ($)', 'Net Profit ($)']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2E75B6', end_color='FF2E75B6', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Row 1 height
    ws.row_dimensions[1].height = 20

    # --- Data rows: A2 = January, A3 = February, A4:A13 = empty ---
    # Only the first two month names are filled; the rest of column A is left blank.
    # Columns B-D have data for all 12 months to make the task meaningful.

    revenues  = [142850, 137240, 158930, 163410, 171250, 184600,
                 192300, 188750, 175420, 166830, 158940, 201300]
    expenses  = [98340,  94120,  103560, 107280, 112440, 119750,
                 125100, 122380, 114260, 108790, 103150, 128450]
    months    = ['January', 'February', 'March', 'April', 'May', 'June',
                 'July', 'August', 'September', 'October', 'November', 'December']

    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    num_fmt = '#,##0.00'

    for i, (rev, exp) in enumerate(zip(revenues, expenses)):
        row = i + 2  # rows 2-13

        # Column A: only fill January (row 2) and February (row 3); leave rest empty
        if row <= 3:
            ws.cell(row=row, column=1, value=months[i]).alignment = Alignment(horizontal='left')

        # Column B: Revenue
        b_cell = ws.cell(row=row, column=2, value=rev)
        b_cell.number_format = num_fmt
        b_cell.border = border

        # Column C: Expenses
        c_cell = ws.cell(row=row, column=3, value=exp)
        c_cell.number_format = num_fmt
        c_cell.border = border

        # Column D: Net Profit (formula-like value, stored as plain number)
        net = rev - exp
        d_cell = ws.cell(row=row, column=4, value=net)
        d_cell.number_format = num_fmt
        d_cell.border = border

        # Alternate row shading
        if i % 2 == 1:
            row_fill = PatternFill(start_color='FFF2F2F2', end_color='FFF2F2F2', fill_type='solid')
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = row_fill

    # Apply borders to column A for all data rows
    for row in range(2, 14):
        cell = ws.cell(row=row, column=1)
        cell.border = border

    # Column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 16

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Add a totals row label in row 14
    total_label = ws.cell(row=14, column=1, value='Total')
    total_label.font = Font(bold=True)
    total_label.alignment = Alignment(horizontal='left')

    ws.cell(row=14, column=2, value=sum(revenues)).number_format = num_fmt
    ws.cell(row=14, column=2).font = Font(bold=True)
    ws.cell(row=14, column=3, value=sum(expenses)).number_format = num_fmt
    ws.cell(row=14, column=3).font = Font(bold=True)
    ws.cell(row=14, column=4, value=sum(revenues) - sum(expenses)).number_format = num_fmt
    ws.cell(row=14, column=4).font = Font(bold=True)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('  Sheet: MonthlyReport')
    print('  A2: January, A3: February, A4:A13: empty')
    print('  Columns B-D: revenue/expenses/net for all 12 months')


create_initial()
