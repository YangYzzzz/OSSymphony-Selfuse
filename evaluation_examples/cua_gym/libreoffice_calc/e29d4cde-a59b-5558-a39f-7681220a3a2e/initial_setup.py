"""
Initial Setup: Supplier comparison database from PDF quotes
Task ID: osworld_multi_apps_doc_pdf_calc_012
Domain: libreoffice_calc (multi-app: PDF + Calc)

Creates:
- /home/user/Desktop/quotes/supplier_apex.pdf
- /home/user/Desktop/quotes/supplier_bolt.pdf
- /home/user/Desktop/quotes/supplier_crest.pdf
- /home/user/Desktop/quote_comparison.ods  (empty template, headers only)

Opens the PDFs in a viewer and the ODS in LibreOffice Calc for the agent.
"""

import os
import shlex
import subprocess
import time

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_doc_pdf_calc_012'
DESKTOP = f'{WORKDIR}/Desktop'
QUOTES_DIR = f'{DESKTOP}/quotes'
ODS_PATH = f'{DESKTOP}/quote_comparison.ods'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_pdfs():
    """Create three supplier quote PDFs using fpdf2."""
    try:
        from fpdf import FPDF
    except ImportError:
        subprocess.run(['pip3', 'install', 'fpdf2'], check=True)
        from fpdf import FPDF

    os.makedirs(QUOTES_DIR, exist_ok=True)

    # Product data: (Product_Code, Product_Name, Min_Order_Qty)
    products = [
        ('PRD-001', 'Industrial Bearing 6205',   50),
        ('PRD-002', 'Hydraulic Seal Kit HS-300', 20),
        ('PRD-003', 'Steel Coupling SC-150',      100),
        ('PRD-004', 'Pneumatic Valve PV-80',      30),
        ('PRD-005', 'Electrical Relay ER-24V',    75),
    ]

    # Supplier data: (supplier_name, short_name, filename, unit_prices)
    suppliers = [
        (
            'Apex Supply Co.',
            'apex',
            'supplier_apex.pdf',
            [12.50, 34.80, 8.75, 67.20, 15.40],  # unit prices per product
        ),
        (
            'Bolt Industries',
            'bolt',
            'supplier_bolt.pdf',
            [11.90, 36.50, 9.20, 63.80, 16.10],
        ),
        (
            'Crest Wholesale',
            'crest',
            'supplier_crest.pdf',
            [13.20, 33.60, 8.50, 65.50, 14.90],
        ),
    ]

    for sup_name, sup_short, filename, prices in suppliers:
        pdf = FPDF()
        pdf.add_page()

        # Header
        pdf.set_font('Helvetica', 'B', 16)
        pdf.cell(0, 12, f'Price Quotation - {sup_name}', ln=True, align='C')
        pdf.set_font('Helvetica', '', 10)
        pdf.cell(0, 8, 'Quote Date: 2026-03-01  |  Valid Until: 2026-04-30', ln=True, align='C')
        pdf.ln(4)

        # Contact / reference block
        pdf.set_font('Helvetica', 'B', 10)
        pdf.cell(0, 7, f'Supplier: {sup_name}', ln=True)
        pdf.set_font('Helvetica', '', 10)
        pdf.cell(0, 7, 'Contact: procurement@' + sup_short + '.com   Tel: +1-800-555-' + str(1000 + ord(sup_short[0])), ln=True)
        pdf.cell(0, 7, 'Payment Terms: Net 30   |   Delivery: 5-7 business days', ln=True)
        pdf.ln(6)

        # Table header
        col_w = [28, 72, 35, 42]
        pdf.set_font('Helvetica', 'B', 10)
        pdf.set_fill_color(200, 220, 255)
        pdf.cell(col_w[0], 8, 'Product Code', border=1, fill=True)
        pdf.cell(col_w[1], 8, 'Product Name', border=1, fill=True)
        pdf.cell(col_w[2], 8, 'Unit Price', border=1, fill=True)
        pdf.cell(col_w[3], 8, 'Min Order Qty', border=1, fill=True, ln=True)

        # Table rows
        pdf.set_font('Helvetica', '', 10)
        for i, (code, name, moq) in enumerate(products):
            pdf.cell(col_w[0], 8, code, border=1)
            pdf.cell(col_w[1], 8, name, border=1)
            pdf.cell(col_w[2], 8, f'${prices[i]:.2f}', border=1)
            pdf.cell(col_w[3], 8, str(moq), border=1, ln=True)

        pdf.ln(6)
        pdf.set_font('Helvetica', 'I', 9)
        pdf.cell(0, 7, 'All prices in USD. Prices subject to change without notice.', ln=True)
        pdf.cell(0, 7, f'Thank you for considering {sup_name} as your supplier.', ln=True)

        out_path = os.path.join(QUOTES_DIR, filename)
        pdf.output(out_path)
        print(f'PDF created: {out_path}')


def create_comparison_ods():
    """Create an empty ODS comparison table with column headers only."""
    try:
        import openpyxl
    except ImportError:
        subprocess.run(['pip3', 'install', 'openpyxl'], check=True)
        import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Quote Comparison'

    # Column headers only — NO data rows (agent must fill from PDFs)
    headers = [
        'Product_Code',
        'Product_Name',
        'Apex_Price',
        'Bolt_Price',
        'Crest_Price',
        'Cheapest_Supplier',
        'Best_Price',
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Set reasonable column widths
    col_widths = [15, 30, 12, 12, 12, 20, 12]
    import string
    for i, width in enumerate(col_widths):
        ws.column_dimensions[string.ascii_uppercase[i]].width = width

    # Save as .ods — use xlrd/openpyxl xlsx then convert, or just save xlsx with .ods name
    # LibreOffice can open xlsx; but task expects .ods. Save as xlsx first, then convert via libreoffice headless.
    xlsx_tmp = f'{WORKDIR}/quote_comparison_tmp.xlsx'
    wb.save(xlsx_tmp)

    # Convert to ODS using LibreOffice headless
    result = subprocess.run(
        ['libreoffice', '--headless', '--convert-to', 'ods',
         '--outdir', DESKTOP, xlsx_tmp],
        capture_output=True, text=True, timeout=60
    )
    if result.returncode != 0:
        print(f'Conversion warning: {result.stderr}')
        # Fallback: just use xlsx as the file
        import shutil
        shutil.copy(xlsx_tmp, ODS_PATH.replace('.ods', '.xlsx'))
        print(f'Fallback: saved as xlsx')
    else:
        # Remove temp file
        if os.path.exists(xlsx_tmp):
            os.remove(xlsx_tmp)
        # libreoffice outputs to DESKTOP with same name stem
        converted = os.path.join(DESKTOP, 'quote_comparison_tmp.ods')
        if os.path.exists(converted):
            import shutil
            shutil.move(converted, ODS_PATH)
    print(f'Comparison table created: {ODS_PATH}')


def main():
    create_pdfs()
    create_comparison_ods()

    # GUI-ready startup: open PDFs in document viewer and Calc file
    time.sleep(1)

    # Open the ODS file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{ODS_PATH}"', delay_sec=2.0)

    # Open PDFs in evince (document viewer) — open them sequentially
    for pdf_name in ['supplier_apex.pdf', 'supplier_bolt.pdf', 'supplier_crest.pdf']:
        pdf_path = os.path.join(QUOTES_DIR, pdf_name)
        launch_gui(f'evince "{pdf_path}"', delay_sec=1.0)

    print('GUI_READY: launched LibreOffice Calc and PDF viewer windows with DISPLAY=:0')


main()
