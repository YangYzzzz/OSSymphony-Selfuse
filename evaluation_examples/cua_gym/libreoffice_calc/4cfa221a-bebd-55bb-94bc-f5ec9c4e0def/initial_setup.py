"""
Initial Setup: Format Paintbrush task - copy formatting from B2 to D2, F2, H2
Task ID: calc_lf_075
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_075'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Dashboard ---
    ws = wb.active
    ws.title = 'Dashboard'

    # Headers
    headers = ['Category', 'Q1 Revenue', 'Status', 'Q2 Revenue', 'Region',
               'Q3 Revenue', 'Notes', 'Q4 Revenue']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='center')

    # Data rows - realistic business data
    data = [
        ['Electronics',  45230.75, 'Active',   15000, 'North America', 22500, 'Strong growth',       31000],
        ['Apparel',      28190.50, 'Active',   19200, 'Europe',        24800, 'Seasonal uptick',     27650],
        ['Home & Garden', 33450.00, 'Review',  21300, 'Asia Pacific',  18900, 'New market entry',    29100],
        ['Automotive',   67820.25, 'Active',   42500, 'North America', 38700, 'Fleet contracts',     51200],
        ['Food & Bev',   19875.60, 'Paused',   12800, 'Europe',        15600, 'Regulatory delay',    18300],
        ['Healthcare',   54310.80, 'Active',   38900, 'Asia Pacific',  41200, 'Expanding coverage',  46800],
        ['Technology',   89450.00, 'Active',   62300, 'North America', 71500, 'Cloud migration',     78200],
        ['Energy',       41200.35, 'Review',   29800, 'Europe',        33400, 'Green initiative',    37600],
        ['Finance',      72690.90, 'Active',   51200, 'Asia Pacific',  58300, 'M&A activity',        64100],
        ['Retail',       26540.15, 'Active',   18700, 'North America', 21900, 'E-commerce push',     24300],
        ['Logistics',    38920.40, 'Active',   27600, 'Europe',        31200, 'Supply chain opt',    35800],
        ['Education',    15780.25, 'Paused',   11200, 'Asia Pacific',  13500, 'Program redesign',    16900],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Apply special formatting to B2 ONLY: blue font, bold, bottom border, '#,##0.00'
    b2 = ws['B2']
    b2.font = Font(color='0000FF', bold=True, size=11)
    bottom_border = Border(bottom=Side(style='thin', color='000000'))
    b2.border = bottom_border
    b2.number_format = '#,##0.00'

    # D2, F2, H2 must remain with DEFAULT formatting (no blue, no bold, no border, no number format)
    # They already have default formatting from cell creation, so nothing to do.

    # Set reasonable column widths
    col_widths = {'A': 16, 'B': 14, 'C': 10, 'D': 14, 'E': 16, 'F': 14, 'G': 20, 'H': 14}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
