"""
Reward Script: Format Paintbrush copy from B2 to D2, F2, H2
Task ID: calc_lf_075
Domain: libreoffice_calc
Scoring:
  - Component 1: D2 has blue bold font, bottom border, #,##0.00 format (0.34 pts)
  - Component 2: F2 has blue bold font, bottom border, #,##0.00 format (0.33 pts)
  - Component 3: H2 has blue bold font, bottom border, #,##0.00 format (0.33 pts)
"""

import os
import time


WORKDIR = '/home/user'
TASK_ID = 'calc_lf_075'


def persist_app_state(domain: str):
    """Save any unsaved edits in the GUI application."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def check_cell_formatting(ws, cell_ref):
    """
    Check if a cell has the expected formatting copied from B2:
      - Bold font
      - Blue font color (000000FF in ARGB)
      - Bottom border (thin)
      - Number format #,##0.00
    Returns (passed: bool, details: str)
    """
    cell = ws[cell_ref]
    results = []
    pass_count = 0

    # Check bold
    is_bold = cell.font.bold is True
    results.append(f"bold={'PASS' if is_bold else 'FAIL'}({cell.font.bold})")
    if is_bold:
        pass_count += 1

    # Check blue font color (000000FF = ARGB for #0000FF blue)
    try:
        fc = cell.font.color.rgb if cell.font.color and cell.font.color.rgb else None
        # Accept both 000000FF and FF0000FF as valid blue representations
        is_blue = fc is not None and isinstance(fc, str) and fc.upper().endswith('0000FF')
        results.append(f"font_color={'PASS' if is_blue else 'FAIL'}({fc})")
        if is_blue:
            pass_count += 1
    except Exception:
        results.append("font_color=FAIL(error)")

    # Check bottom border
    has_bottom = cell.border.bottom.style is not None
    results.append(f"bottom_border={'PASS' if has_bottom else 'FAIL'}({cell.border.bottom.style})")
    if has_bottom:
        pass_count += 1

    # Check number format
    has_nf = cell.number_format == '#,##0.00'
    results.append(f"num_format={'PASS' if has_nf else 'FAIL'}({cell.number_format})")
    if has_nf:
        pass_count += 1

    all_pass = (pass_count == 4)

    return all_pass, '; '.join(results)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify 'Dashboard' sheet exists
    if 'Dashboard' not in wb.sheetnames:
        print(f"CRITICAL: 'Dashboard' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Dashboard']

    # Component 1: D2 formatting matches B2 (0.34 points)
    try:
        passed, details = check_cell_formatting(ws, 'D2')
        if passed:
            print(f"PASS: Component 1 - D2 formatting copied from B2 (0.34 pts) [{details}]")
            total_score += 0.34
        else:
            print(f"FAIL: Component 1 - D2 formatting incomplete [{details}]")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: F2 formatting matches B2 (0.33 points)
    try:
        passed, details = check_cell_formatting(ws, 'F2')
        if passed:
            print(f"PASS: Component 2 - F2 formatting copied from B2 (0.33 pts) [{details}]")
            total_score += 0.33
        else:
            print(f"FAIL: Component 2 - F2 formatting incomplete [{details}]")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: H2 formatting matches B2 (0.33 points)
    try:
        passed, details = check_cell_formatting(ws, 'H2')
        if passed:
            print(f"PASS: Component 3 - H2 formatting copied from B2 (0.33 pts) [{details}]")
            total_score += 0.33
        else:
            print(f"FAIL: Component 3 - H2 formatting incomplete [{details}]")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist any unsaved GUI state before verification
persist_app_state("libreoffice_calc")

# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
