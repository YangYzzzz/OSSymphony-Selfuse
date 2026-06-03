"""
Initial Setup: Sensor data tracking spreadsheet with some missing Reading values.
Task ID: osworld_multi_apps_calc_vscode_009
Domain: libreoffice_calc (multi-app: Calc + VSCode)

Creates sensor_data.xlsx on the Desktop with hourly sensor readings and
some empty Reading cells. Opens the file in LibreOffice Calc.
Also opens VSCode for the Python scripting portion of the task.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user/Desktop'
TASK_ID = 'sensor_data'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    os.makedirs(WORKDIR, exist_ok=True)

    wb = openpyxl.Workbook()

    # --- Sheet: SensorReadings ---
    ws = wb.active
    ws.title = 'SensorReadings'

    # Header row styling
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2E75B6', end_color='FF2E75B6', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Headers
    headers = ['Timestamp', 'Sensor_ID', 'Reading']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # Hourly sensor data for 2025-03-15 with some missing Reading values (None)
    # Missing at indices: 1, 4, 9, 13, 17, 21 (0-indexed)
    sensor_readings = [
        23.4, None, 22.8, 23.1, None, 24.5,
        25.2, 25.8, 26.3, None, 27.1, 26.9,
        26.4, None, 25.7, 25.3, 24.9, None,
        24.1, 23.8, 23.5, None, 22.9, 22.6,
    ]

    timestamps = [
        '2025-03-15 00:00', '2025-03-15 01:00', '2025-03-15 02:00',
        '2025-03-15 03:00', '2025-03-15 04:00', '2025-03-15 05:00',
        '2025-03-15 06:00', '2025-03-15 07:00', '2025-03-15 08:00',
        '2025-03-15 09:00', '2025-03-15 10:00', '2025-03-15 11:00',
        '2025-03-15 12:00', '2025-03-15 13:00', '2025-03-15 14:00',
        '2025-03-15 15:00', '2025-03-15 16:00', '2025-03-15 17:00',
        '2025-03-15 18:00', '2025-03-15 19:00', '2025-03-15 20:00',
        '2025-03-15 21:00', '2025-03-15 22:00', '2025-03-15 23:00',
    ]

    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    data_align = Alignment(horizontal='center', vertical='center')

    for row_idx, (ts, reading) in enumerate(zip(timestamps, sensor_readings), 2):
        # Timestamp
        c_ts = ws.cell(row=row_idx, column=1, value=ts)
        c_ts.alignment = data_align
        c_ts.border = data_border

        # Sensor ID
        c_sid = ws.cell(row=row_idx, column=2, value='SENS_A1')
        c_sid.alignment = data_align
        c_sid.border = data_border

        # Reading (may be None/empty for some rows)
        c_r = ws.cell(row=row_idx, column=3, value=reading)
        c_r.alignment = data_align
        c_r.border = data_border
        if reading is not None:
            c_r.number_format = '0.0'

    # Column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open sensor_data.xlsx in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=3.0)

    # Also open VSCode for Python scripting (task requires writing a Python script)
    launch_gui('code "/home/user/Desktop"', delay_sec=2.0)

    print('GUI_READY: launched LibreOffice Calc and VSCode with DISPLAY=:0')


create_initial()
