"""
Reward Script: HR Probation Tracker — Automated formulas and conditional formatting
Task ID: calc_hr_probation_tracker_005
Domain: libreoffice_calc

Scoring Rubric:
  Component 1 (0.35): D2:D23 contain =C{row}+90 formula for probation end date
  Component 2 (0.35): E2:E23 contain IF(TODAY()<=D{row},...) formula for status
  Component 3 (0.30): Conditional formatting on A2:F23 — red for Overdue, green for Active Probation
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_probation_tracker_005'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the sheet exists
    if 'Onboarding' not in wb.sheetnames:
        print("FAIL: 'Onboarding' sheet not found")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws = wb['Onboarding']

    # Component 1: D2:D23 contain =C{row}+90 formulas for probation end date (0.35 points)
    # This verifies that D column (initially empty) now has date formulas
    try:
        d_formula_count = 0
        d_formula_correct = 0
        expected_rows = range(2, 24)  # rows 2-23

        for row in expected_rows:
            cell = ws.cell(row=row, column=4)  # Column D
            val = cell.value
            if val is not None and isinstance(val, str) and val.strip():
                d_formula_count += 1
                # Check the formula matches =C{row}+90 pattern (case-insensitive)
                normalized = val.strip().upper().replace(' ', '')
                expected_formula = f'=C{row}+90'
                if normalized == expected_formula.upper():
                    d_formula_correct += 1

        if d_formula_count == 22 and d_formula_correct == 22:
            print(f"PASS: Component 1 — All 22 D column formulas correct (=C{{row}}+90) (0.35 pts)")
            total_score += 0.35
        elif d_formula_correct >= 11:
            # Partial credit if more than half are correct
            partial = round(0.35 * (d_formula_correct / 22), 4)
            print(f"PARTIAL: Component 1 — {d_formula_correct}/22 D column formulas correct, partial score ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Expected 22 =C{{row}}+90 formulas in D2:D23, found {d_formula_count} with values, {d_formula_correct} correct")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: E2:E23 contain IF(TODAY()<=D{row},...) formulas for status (0.35 points)
    # This verifies that E column (initially empty) now has status formulas using TODAY()
    try:
        e_formula_count = 0
        e_formula_correct = 0

        for row in range(2, 24):
            cell = ws.cell(row=row, column=5)  # Column E
            val = cell.value
            if val is not None and isinstance(val, str) and val.strip():
                e_formula_count += 1
                # Check the formula contains TODAY() and relevant keywords
                val_upper = val.strip().upper().replace(' ', '')
                # Must have IF, TODAY, the probation end date reference, and status strings
                has_if = 'IF(' in val_upper
                has_today = 'TODAY()' in val_upper
                has_d_ref = f'D{row}' in val_upper
                has_active = 'ACTIVEPROBATION' in val_upper.replace('"', '').replace("'", '')
                has_overdue = 'OVERDUE' in val_upper.replace('"', '').replace("'", '')

                if has_if and has_today and has_d_ref and has_active and has_overdue:
                    e_formula_correct += 1

        if e_formula_count == 22 and e_formula_correct == 22:
            print(f"PASS: Component 2 — All 22 E column status formulas correct (IF/TODAY) (0.35 pts)")
            total_score += 0.35
        elif e_formula_correct >= 11:
            partial = round(0.35 * (e_formula_correct / 22), 4)
            print(f"PARTIAL: Component 2 — {e_formula_correct}/22 E column formulas correct, partial score ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Expected 22 IF(TODAY()<=D{{row}},...) formulas in E2:E23, found {e_formula_count} with values, {e_formula_correct} correct")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Conditional formatting on A2:F23 with red for Overdue and green for Active Probation (0.30 points)
    # This verifies conditional formatting was added (initially none)
    try:
        cf_rules = ws.conditional_formatting
        cf_list = list(cf_rules)

        has_overdue_red = False
        has_active_green = False

        for cf_range in cf_list:
            rules = ws.conditional_formatting[cf_range]
            for rule in rules:
                if rule.type in ('expression', 'formula'):
                    formula_str = ' '.join(str(f) for f in rule.formula).upper() if rule.formula else ''
                    # Check for Overdue rule with red background (FF0000)
                    if 'OVERDUE' in formula_str and rule.dxf and rule.dxf.fill:
                        try:
                            fill_color = rule.dxf.fill.fgColor.rgb
                            overdue_color_ok = (fill_color is not None and 'FF0000' in fill_color.upper())
                            if overdue_color_ok:
                                has_overdue_red = overdue_color_ok
                        except Exception:
                            pass
                    # Check for Active Probation rule with green background (70AD47)
                    if 'ACTIVEPROBATION' in formula_str.replace(' ', '').replace('"', '').replace("'", '') and rule.dxf and rule.dxf.fill:
                        try:
                            fill_color = rule.dxf.fill.fgColor.rgb
                            active_color_ok = (fill_color is not None and '70AD47' in fill_color.upper())
                            if active_color_ok:
                                has_active_green = active_color_ok
                        except Exception:
                            pass

        if has_overdue_red and has_active_green:
            print(f"PASS: Component 3 — Both conditional formatting rules present (red for Overdue, green for Active Probation) (0.30 pts)")
            total_score += 0.30
        elif has_overdue_red or has_active_green:
            found = []
            if has_overdue_red:
                found.append("Overdue=red")
            if has_active_green:
                found.append("Active Probation=green")
            print(f"PARTIAL: Component 3 — Only {found} found, missing the other rule (0.15 pts)")
            total_score += 0.15
        elif len(cf_list) > 0:
            print(f"PARTIAL: Component 3 — Conditional formatting exists ({len(cf_list)} range(s)) but rules don't match expected colors/formulas (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — No conditional formatting found on worksheet")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
