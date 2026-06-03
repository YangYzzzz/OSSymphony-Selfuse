"""
Initial Setup: HR Probation Tracker - New Hire Onboarding Spreadsheet
Task ID: calc_hr_probation_tracker_005
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_probation_tracker_005'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Onboarding ---
    ws = wb.active
    ws.title = 'Onboarding'

    # Headers
    headers = ['Emp ID', 'Name', 'Start Date', 'Probation End', 'Status', 'Supervisor']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # New hire data — 22 employees hired in last 6 months
    # Using a reference date of 2026-03-04; start dates span from ~Sep 2025 to Feb 2026
    employees = [
        ('EMP-2101', 'Aisha Okonkwo',      date(2025, 9, 10),  'Jennifer Walsh'),
        ('EMP-2102', 'Marcus Delgado',     date(2025, 9, 22),  'Daniel Cho'),
        ('EMP-2103', 'Priya Nair',         date(2025, 10, 5),  'Sandra Patel'),
        ('EMP-2104', 'Tyler Hawkins',      date(2025, 10, 14), 'Jennifer Walsh'),
        ('EMP-2105', 'Fatima Al-Hassan',   date(2025, 10, 27), 'Robert Nguyen'),
        ('EMP-2106', 'Lucas Fernandez',    date(2025, 11, 3),  'Sandra Patel'),
        ('EMP-2107', 'Yuki Tanaka',        date(2025, 11, 12), 'Daniel Cho'),
        ('EMP-2108', 'Brianna Scott',      date(2025, 11, 18), 'Robert Nguyen'),
        ('EMP-2109', 'Omar Hassan',        date(2025, 12, 1),  'Jennifer Walsh'),
        ('EMP-2110', 'Chloe Beaumont',     date(2025, 12, 9),  'Sandra Patel'),
        ('EMP-2111', 'Kevin Osei',         date(2025, 12, 15), 'Daniel Cho'),
        ('EMP-2112', 'Natasha Ivanova',    date(2025, 12, 22), 'Robert Nguyen'),
        ('EMP-2113', 'Diego Ramirez',      date(2026, 1, 6),   'Jennifer Walsh'),
        ('EMP-2114', 'Hannah Lee',         date(2026, 1, 13),  'Sandra Patel'),
        ('EMP-2115', 'Jamal Williams',     date(2026, 1, 20),  'Daniel Cho'),
        ('EMP-2116', 'Sofia Moretti',      date(2026, 1, 27),  'Robert Nguyen'),
        ('EMP-2117', 'Ethan Park',         date(2026, 2, 3),   'Jennifer Walsh'),
        ('EMP-2118', 'Amara Diallo',       date(2026, 2, 10),  'Sandra Patel'),
        ('EMP-2119', 'Connor Murphy',      date(2026, 2, 17),  'Daniel Cho'),
        ('EMP-2120', 'Lena Hofmann',       date(2026, 2, 24),  'Robert Nguyen'),
        ('EMP-2121', 'Ravi Krishnamurthy', date(2026, 3, 1),   'Jennifer Walsh'),
        ('EMP-2122', 'Isabelle Dupont',    date(2026, 3, 3),   'Sandra Patel'),
    ]

    for row_idx, (emp_id, name, start_date, supervisor) in enumerate(employees, 2):
        ws.cell(row=row_idx, column=1, value=emp_id)
        ws.cell(row=row_idx, column=2, value=name)
        # Start Date as date value
        start_cell = ws.cell(row=row_idx, column=3, value=start_date)
        start_cell.number_format = 'YYYY-MM-DD'
        # Column D (Probation End) — EMPTY, task will fill
        ws.cell(row=row_idx, column=4, value=None)
        # Column E (Status) — EMPTY, task will fill
        ws.cell(row=row_idx, column=5, value=None)
        ws.cell(row=row_idx, column=6, value=supervisor)

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 20

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

create_initial()
