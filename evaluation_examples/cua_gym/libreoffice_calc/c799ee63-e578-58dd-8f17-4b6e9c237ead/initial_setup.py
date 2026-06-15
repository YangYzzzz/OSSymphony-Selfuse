"""
Initial Setup: Add diagonal border to cells A3:E3 in a matrix form spreadsheet
Task ID: calc_gfl_074
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_074'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Form"

    # Header style
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # Column headers in row 1
    headers = ["Parameter A", "Parameter B", "Parameter C", "Parameter D", "Parameter E"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Set column widths
    for col_letter in ["A", "B", "C", "D", "E"]:
        ws.column_dimensions[col_letter].width = 16

    # Matrix data - 20 rows of configuration combinations
    # Row 3 will be the "impossible combination" row (empty or marked N/A conceptually)
    matrix_data = [
        # Row 2
        [12.5, 0.80, 340, 1.02, 55.0],
        # Row 3 - logically impossible combination (cells left empty, to be marked with diagonal)
        [None, None, None, None, None],
        # Row 4
        [15.0, 0.92, 410, 1.15, 62.3],
        # Row 5
        [18.2, 1.05, 285, 0.98, 48.7],
        # Row 6
        [11.8, 0.73, 520, 1.30, 71.2],
        # Row 7
        [20.1, 1.12, 395, 1.08, 59.4],
        # Row 8
        [14.6, 0.88, 460, 1.22, 66.8],
        # Row 9
        [16.9, 0.95, 310, 1.00, 52.1],
        # Row 10
        [22.3, 1.18, 375, 1.11, 57.6],
        # Row 11
        [13.4, 0.82, 490, 1.27, 69.5],
        # Row 12
        [17.7, 1.01, 350, 1.05, 54.3],
        # Row 13
        [19.5, 1.09, 425, 1.18, 63.9],
        # Row 14
        [10.2, 0.70, 540, 1.35, 73.8],
        # Row 15
        [21.8, 1.15, 295, 0.96, 47.2],
        # Row 16
        [15.3, 0.90, 480, 1.25, 68.1],
        # Row 17
        [18.6, 1.03, 365, 1.07, 56.0],
        # Row 18
        [12.1, 0.77, 505, 1.32, 72.4],
        # Row 19
        [23.0, 1.20, 330, 1.03, 53.7],
        # Row 20
        [16.4, 0.93, 445, 1.20, 65.5],
        # Row 21
        [14.0, 0.85, 385, 1.13, 60.8],
    ]

    # Data cell style
    data_align = Alignment(horizontal="center", vertical="center")
    data_font = Font(name="Arial", size=11)

    for r_idx, row_data in enumerate(matrix_data, 2):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = data_align
            cell.font = data_font
            cell.border = thin_border
            if val is not None:
                cell.number_format = '0.00'

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
