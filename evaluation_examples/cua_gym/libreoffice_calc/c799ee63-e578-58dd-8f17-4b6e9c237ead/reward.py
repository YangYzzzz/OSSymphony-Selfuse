"""
Reward Script: Add inner diagonal border (top-left to bottom-right) to cells A3:E3
Task ID: calc_gfl_074
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): All 5 cells A3:E3 have diagonalDown=True with thin style
  Component 2 (0.3): Diagonal applied precisely — no spurious diagonals + diagonalDown on all 5
  Component 3 (0.2): Cell content unchanged AND diagonal present (compound check)
"""

import openpyxl
import os

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_074'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Form' sheet must exist
    if 'Form' not in wb.sheetnames:
        print(f"FAIL: Sheet 'Form' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Form']
    target_cells = ['A3', 'B3', 'C3', 'D3', 'E3']

    # Component 1: All 5 cells A3:E3 have diagonalDown=True with thin diagonal style (0.5 points)
    # This is the core task requirement — diagonal border with correct direction and style
    try:
        full_match_count = 0
        for coord in target_cells:
            cell = ws[coord]
            b = cell.border
            diag_style = b.diagonal.style if b.diagonal else None
            has_diag_down = b.diagonalDown
            if has_diag_down and diag_style == 'thin':
                full_match_count += 1
            else:
                print(f"FAIL: {coord} — diagonalDown={has_diag_down}, diag_style={diag_style} (expected True, 'thin')")

        if full_match_count == 5:
            print(f"PASS: Component 1 — All 5 cells have diagonalDown=True with thin style (0.5 pts)")
            total_score += 0.5
        elif full_match_count > 0:
            partial = round(0.5 * (full_match_count / 5), 2)
            print(f"PARTIAL: Component 1 — {full_match_count}/5 cells fully match ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No cells have the required diagonal border")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Diagonal applied precisely — all 5 target cells have it AND no other cells do (0.3 points)
    # This is a compound check: anchored on the task-introduced diagonal being present
    try:
        # First check that all target cells have diagonal (gate condition)
        diag_target_count = sum(
            1 for coord in target_cells
            if ws[coord].border.diagonalDown and (ws[coord].border.diagonal.style if ws[coord].border.diagonal else None) is not None
        )

        if diag_target_count < len(target_cells):
            print(f"FAIL: Component 2 — Not all target cells have diagonal border; precision check skipped")
        else:
            # Now check no spurious diagonals
            spurious_diag = []
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.coordinate in target_cells:
                        continue
                    b = cell.border
                    if b.diagonal and b.diagonal.style:
                        spurious_diag.append(cell.coordinate)

            if len(spurious_diag) == 0:
                print(f"PASS: Component 2 — Diagonal applied precisely to A3:E3 only (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — Spurious diagonal borders at: {spurious_diag}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Cell content in A3:E3 unchanged AND diagonal present (0.2 points)
    # Compound check: the diagonal was added without modifying cell content
    try:
        # Gate: diagonal must be present on all target cells
        diag_count = sum(1 for coord in target_cells if ws[coord].border.diagonalDown)

        if diag_count < len(target_cells):
            print(f"FAIL: Component 3 — Diagonal not present on all cells; content check skipped")
        else:
            modified_cells = [coord for coord in target_cells if ws[coord].value is not None]

            if len(modified_cells) == 0:
                print(f"PASS: Component 3 — Cell content unchanged and diagonal present (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 3 — Cell content was modified in: {modified_cells}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
