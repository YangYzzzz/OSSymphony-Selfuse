"""
Reward Script: Theory Workshop Invitees - Email, Country, and Sort
Task ID: osworld_multi_apps_web_prof_email_015
Domain: libreoffice_calc
Scoring:
  Component 1: Email column populated for all 9 professors (0.35 pts)
  Component 2: Country column populated for all 9 professors (0.30 pts)
  Component 3: Rows sorted alphabetically by Country column (0.35 pts)
Total: 1.0

Task: Open each professor's webpage in Chrome, collect their email and
primary country of affiliation, populate the spreadsheet, sort by Country
alphabetically, and save.

Initial state: Email and Country columns are blank (None).
Golden state: Email and Country filled in, rows sorted by Country A-Z.
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_web_prof_email_015'
FILE_PATH = f'{WORKDIR}/Theory_Workshop_Invitees.xlsx'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook — precondition gate
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Access the active sheet (named 'Invitees')
    try:
        if 'Invitees' in wb.sheetnames:
            ws = wb['Invitees']
        else:
            ws = wb.active
    except Exception as e:
        print(f"CRITICAL: Cannot access worksheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Read all data rows (skip header row 1)
    # Columns: A=Name, B=Webpage, C=Email, D=Country
    try:
        names = []
        emails = []
        countries = []
        for row in ws.iter_rows(min_row=2, max_row=10, values_only=True):
            name, webpage, email, country = row[0], row[1], row[2], row[3]
            names.append(name)
            emails.append(email)
            countries.append(country)
        print(f"INFO: Read {len(names)} data rows")
        print(f"INFO: Names: {names}")
        print(f"INFO: Emails: {emails}")
        print(f"INFO: Countries: {countries}")
    except Exception as e:
        print(f"CRITICAL: Cannot read data rows: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: verify we have 9 data rows
    if len(names) != 9:
        print(f"WARN: Expected 9 data rows, found {len(names)}")

    # Component 1: Email column populated for all 9 professors (0.35 pts)
    # In the initial state, all emails are None.
    # In the golden state, all 9 emails are filled with real addresses.
    try:
        non_null_emails = [e for e in emails if e is not None and str(e).strip() != '' and str(e).strip().lower() != 'none']
        email_count = len(non_null_emails)
        if email_count == 9:
            print(f"PASS: Component 1 — All 9 email addresses populated (0.35 pts)")
            print(f"      Emails: {non_null_emails}")
            total_score += 0.35
        elif email_count > 0:
            # Partial credit based on proportion filled
            partial = round(0.35 * email_count / 9, 4)
            print(f"PARTIAL: Component 1 — {email_count}/9 email addresses populated ({partial} pts)")
            if partial > 0:
                total_score += partial
        else:
            print(f"FAIL: Component 1 — No emails populated (found: {emails})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Country column populated for all 9 professors (0.30 pts)
    # In the initial state, all countries are None.
    # In the golden state, all 9 countries are filled (Canada, Israel, Switzerland, UK, US).
    try:
        non_null_countries = [c for c in countries if c is not None and str(c).strip() != '' and str(c).strip().lower() != 'none']
        country_count = len(non_null_countries)
        if country_count == 9:
            print(f"PASS: Component 2 — All 9 country values populated (0.30 pts)")
            print(f"      Countries: {non_null_countries}")
            total_score += 0.30
        elif country_count > 0:
            partial = round(0.30 * country_count / 9, 4)
            print(f"PARTIAL: Component 2 — {country_count}/9 country values populated ({partial} pts)")
            if partial > 0:
                total_score += partial
        else:
            print(f"FAIL: Component 2 — No countries populated (found: {countries})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Rows sorted alphabetically by Country (0.35 pts)
    # In the initial state, countries are all None so there is no meaningful sort.
    # In the golden state, rows are sorted A-Z by Country:
    #   Canada, Canada, Israel, Switzerland, UK, UK, US, US, US
    # This component also requires that countries are non-null (anchored to the task change).
    try:
        # Only evaluate sort if countries were populated (component 2 passed or partially passed)
        non_null_c = [c for c in countries if c is not None and str(c).strip() != '' and str(c).strip().lower() != 'none']
        if len(non_null_c) == 9:
            country_strings = [str(c).strip() for c in countries]
            sorted_countries = sorted(country_strings, key=lambda x: x.lower())
            actual_lower = [x.lower() for x in country_strings]
            sorted_lower = [x.lower() for x in sorted_countries]
            if actual_lower == sorted_lower:
                print(f"PASS: Component 3 — Rows sorted alphabetically by Country (0.35 pts)")
                print(f"      Country order: {country_strings}")
                total_score += 0.35
            else:
                print(f"FAIL: Component 3 — Rows NOT sorted by Country")
                print(f"      Actual order:   {country_strings}")
                print(f"      Expected order: {sorted_countries}")
        else:
            print(f"FAIL: Component 3 — Cannot check sort order; only {len(non_null_c)}/9 countries populated")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


if not os.path.exists(FILE_PATH):
    print(f"File not found: {FILE_PATH}")
    print("REWARD: 0.0")
else:
    verify_task(FILE_PATH)
