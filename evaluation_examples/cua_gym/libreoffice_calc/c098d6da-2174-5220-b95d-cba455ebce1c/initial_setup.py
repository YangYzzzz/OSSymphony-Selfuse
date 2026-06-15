"""
Initial Setup: Theory Workshop Invitees spreadsheet - professors' names and webpage URLs
Task ID: osworld_multi_apps_web_prof_email_015
Domain: libreoffice_calc (multi-app: also Chrome)
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_web_prof_email_015'
OUTPUT = f'{WORKDIR}/Theory_Workshop_Invitees.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Invitees ---
    ws = wb.active
    ws.title = 'Invitees'

    # Headers
    headers = ['Name', 'Webpage', 'Email', 'Country']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 9 professors spanning US (3), UK (2), Canada (2), Israel (1), Switzerland (1)
    # Email and Country columns are intentionally BLANK (agent must fill by visiting webpages)
    # Listed in unsorted order (not sorted by country) — agent must sort after filling
    professors = [
        # Name, Webpage
        # US professors (3)
        ['Scott Aaronson',       'https://www.cs.utexas.edu/~aaronson/'],
        ['Salil Vadhan',         'https://salil.seas.harvard.edu/'],
        ['Tim Roughgarden',      'https://timroughgarden.org/'],
        # UK professors (2)
        ['Leslie Ann Goldberg',  'https://www.cs.ox.ac.uk/people/leslie.goldberg/'],
        ['Mark Jerrum',          'https://www.maths.qmul.ac.uk/~mj/'],
        # Canada professors (2)
        ['Toniann Pitassi',      'https://www.cs.toronto.edu/~toni/'],
        ['Eric Blais',           'https://cs.uwaterloo.ca/~eblais/'],
        # Israel professors (1)
        ['Oded Goldreich',       'https://www.wisdom.weizmann.ac.il/~oded/'],
        # Switzerland professors (1)
        ['Thomas Holenstein',    'https://inf.ethz.ch/people/person-detail.holenstein.html'],
    ]

    for r, row_data in enumerate(professors, 2):
        ws.cell(row=r, column=1, value=row_data[0])  # Name
        ws.cell(row=r, column=2, value=row_data[1])  # Webpage
        # Email (column 3) left blank — agent must fill from webpages
        ws.cell(row=r, column=3, value='')
        # Country (column 4) left blank — agent must fill from webpages
        ws.cell(row=r, column=4, value='')

    # Set column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open LibreOffice Calc with the file, then open Chrome
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    launch_gui('google-chrome', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc and Chrome with DISPLAY=:0')


create_initial()
