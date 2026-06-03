"""
Initial Setup: Create spreadsheet with formatted Sheet1 and unformatted Sheet2
Task ID: calc_tbl_034
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_034'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

# -- Style definitions --
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

LIGHT_ROW_FILL = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")
DARK_ROW_FILL = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)

DATA_FONT = Font(name="Calibri", size=11)


def launch_gui(command: str, delay_sec: float = 1.0):
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # -- Headers --
    headers = ["Employee ID", "Full Name", "Department", "Hire Date", "Annual Salary", "Performance Score"]

    # -- Data rows (19 rows of realistic content) --
    data = [
        ["EMP-1001", "Sarah Chen", "Engineering", "2021-03-15", 92500, 4.3],
        ["EMP-1002", "Marcus Johnson", "Marketing", "2020-07-22", 78000, 3.8],
        ["EMP-1003", "Priya Patel", "Finance", "2019-11-03", 88000, 4.1],
        ["EMP-1004", "James O'Brien", "Engineering", "2022-01-10", 85000, 3.5],
        ["EMP-1005", "Aisha Mohammed", "Human Resources", "2018-06-28", 72000, 4.6],
        ["EMP-1006", "David Kim", "Engineering", "2023-02-14", 95000, 4.0],
        ["EMP-1007", "Elena Rodriguez", "Sales", "2020-09-05", 68500, 3.9],
        ["EMP-1008", "Thomas Wright", "Finance", "2021-04-19", 83000, 4.2],
        ["EMP-1009", "Fatima Al-Hassan", "Marketing", "2022-08-30", 74500, 3.7],
        ["EMP-1010", "Robert Tanaka", "Operations", "2019-12-01", 79000, 4.4],
        ["EMP-1011", "Lisa Nguyen", "Engineering", "2023-05-16", 91000, 3.6],
        ["EMP-1012", "Carlos Mendoza", "Sales", "2020-02-28", 71000, 4.0],
        ["EMP-1013", "Hannah Becker", "Human Resources", "2021-10-11", 69500, 4.5],
        ["EMP-1014", "Omar Farouk", "Finance", "2022-06-07", 86500, 3.8],
        ["EMP-1015", "Jennifer Walsh", "Operations", "2018-03-25", 77000, 4.1],
        ["EMP-1016", "Wei Zhang", "Engineering", "2023-09-01", 98000, 4.7],
        ["EMP-1017", "Rachel Foster", "Marketing", "2019-05-14", 73000, 3.4],
        ["EMP-1018", "Dmitri Volkov", "Sales", "2021-12-20", 67500, 3.9],
        ["EMP-1019", "Sophie Martin", "Finance", "2020-04-08", 84000, 4.3],
    ]

    # ===== Sheet1: Formatted data =====
    ws1 = wb.active
    ws1.title = "Sheet1"

    # Write headers with formatting
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Write data with alternating row colors and borders
    for r, row_data in enumerate(data, 2):
        fill = LIGHT_ROW_FILL if r % 2 == 0 else DARK_ROW_FILL
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER

    # Set column widths for readability
    ws1.column_dimensions["A"].width = 14
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 14
    ws1.column_dimensions["E"].width = 16
    ws1.column_dimensions["F"].width = 20

    # ===== Sheet2: Same data, NO formatting =====
    ws2 = wb.create_sheet("Sheet2")

    # Write headers (plain, no formatting)
    for col, h in enumerate(headers, 1):
        ws2.cell(row=1, column=col, value=h)

    # Write data (plain, no formatting)
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
