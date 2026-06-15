"""
Reward Script: Paste Special (formats only) from Sheet1 to Sheet2
Task ID: calc_tbl_034
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.35): Header row bold formatting on Sheet2 A1:F1
  - Component 2 (0.25): Header row background fill (FF2F5496) on Sheet2 A1:F1
  - Component 3 (0.25): Alternating row colors on Sheet2 rows 2-20
  - Component 4 (0.15): Thin borders on Sheet2 A1:F20
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_034'


def persist_app_state(domain: str):
    """Best-effort save via Ctrl+S for unsaved GUI edits."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            import time
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify that Sheet2 A1:F20 has formatting copied from Sheet1.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sheet2 must exist
    if 'Sheet2' not in wb.sheetnames:
        print("FAIL: Sheet2 not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws2 = wb['Sheet2']

    # Component 1: Header row (row 1) has bold formatting on A1:F1 (0.35 points)
    # In initial_env, row 1 is NOT bold. In golden_env, row 1 IS bold.
    try:
        bold_count = 0
        total_header_cells = 6  # A1:F1
        for c in range(1, 7):
            cell = ws2.cell(row=1, column=c)
            if cell.font.bold:
                bold_count += 1
        if bold_count == total_header_cells:
            print(f"PASS: Component 1 — All 6 header cells are bold ({bold_count}/6) (0.35 pts)")
            total_score += 0.35
        elif bold_count > 0:
            partial = 0.35 * (bold_count / total_header_cells)
            print(f"PARTIAL: Component 1 — {bold_count}/6 header cells are bold ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No header cells are bold (0/6)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Header row background fill is FF2F5496 (dark blue) (0.25 points)
    # In initial_env, fill is 00000000 (none). In golden_env, fill is FF2F5496.
    try:
        fill_count = 0
        for c in range(1, 7):
            cell = ws2.cell(row=1, column=c)
            try:
                fill_rgb = cell.fill.fgColor.rgb
                if fill_rgb == 'FF2F5496':
                    fill_count += 1
            except Exception:
                pass
        if fill_count == total_header_cells:
            print(f"PASS: Component 2 — All 6 header cells have correct fill color FF2F5496 ({fill_count}/6) (0.25 pts)")
            total_score += 0.25
        elif fill_count > 0:
            partial = 0.25 * (fill_count / total_header_cells)
            print(f"PARTIAL: Component 2 — {fill_count}/6 header cells have correct fill ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No header cells have fill FF2F5496 (0/6)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Alternating row colors on rows 2-20 (0.25 points)
    # Even rows (2,4,6,...,20) should have fill FFD6E4F0 (light blue)
    # Odd rows (3,5,7,...,19) should have fill FFFFFFFF (white)
    # In initial_env, all rows have fill 00000000.
    try:
        correct_alt_count = 0
        total_data_rows = 19  # rows 2-20
        for r in range(2, 21):
            cell = ws2.cell(row=r, column=1)
            try:
                fill_rgb = cell.fill.fgColor.rgb
            except Exception:
                fill_rgb = None

            if r % 2 == 0:
                # Even rows should be light blue
                if fill_rgb == 'FFD6E4F0':
                    correct_alt_count += 1
            else:
                # Odd rows should be white
                if fill_rgb == 'FFFFFFFF':
                    correct_alt_count += 1

        if correct_alt_count == total_data_rows:
            print(f"PASS: Component 3 — All 19 data rows have correct alternating colors ({correct_alt_count}/19) (0.25 pts)")
            total_score += 0.25
        elif correct_alt_count > 0:
            partial = 0.25 * (correct_alt_count / total_data_rows)
            print(f"PARTIAL: Component 3 — {correct_alt_count}/19 rows have correct alternating colors ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No data rows have correct alternating colors (0/19)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Thin borders on Sheet2 A1:F20 (0.15 points)
    # In initial_env, no borders. In golden_env, thin borders on all cells.
    try:
        border_count = 0
        total_cells = 20 * 6  # 20 rows x 6 columns = 120 cells
        for r in range(1, 21):
            for c in range(1, 7):
                cell = ws2.cell(row=r, column=c)
                # Check if at least left border is thin
                if cell.border.left.style == 'thin':
                    border_count += 1
        if border_count == total_cells:
            print(f"PASS: Component 4 — All 120 cells have thin borders ({border_count}/120) (0.15 pts)")
            total_score += 0.15
        elif border_count >= total_cells * 0.8:
            partial = 0.15 * (border_count / total_cells)
            print(f"PARTIAL: Component 4 — {border_count}/120 cells have thin borders ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {border_count}/120 cells have thin borders")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
