"""
Initial Setup: ArXiv cs.CL papers deduplication tracker
Task ID: osworld_multi_apps_arxiv_llms_calc_011
Domain: libreoffice_calc

Creates dedup_tracker.ods with empty paper data structure (headers only),
and a Summary section layout. Chrome and LibreOffice Calc are opened.
"""

import os
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_arxiv_llms_calc_011'
XLSX_TMP = f'{WORKDIR}/{TASK_ID}_tmp.xlsx'
OUTPUT   = f'{WORKDIR}/dedup_tracker.ods'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env['DISPLAY'] = ':0'
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ---------------------------------------------------------------
    # Sheet1: Paper data table
    # ---------------------------------------------------------------
    ws1 = wb.active
    ws1.title = 'Sheet1'

    # Main table headers (columns A-E)
    headers = ['arXiv ID', 'Title', 'Authors', 'Date', 'Status']
    header_font = Font(bold=True)
    for col_idx, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_idx, value=h)
        cell.font = header_font

    # Summary section headers (columns G-H, row 1)
    ws1.cell(row=1, column=7, value='Date').font = header_font
    ws1.cell(row=1, column=8, value='Count').font = header_font

    # Summary date placeholders (rows 2-4, columns G-H)
    # These rows are present but empty — agent fills in counts
    summary_dates = ['2024-01-15', '2024-01-16', '2024-01-17']
    for row_offset, d in enumerate(summary_dates, 2):
        ws1.cell(row=row_offset, column=7, value=d)
        # Count column is empty (to be filled by COUNTIF)

    # Column widths for readability
    ws1.column_dimensions['A'].width = 18   # arXiv ID
    ws1.column_dimensions['B'].width = 55   # Title
    ws1.column_dimensions['C'].width = 40   # Authors
    ws1.column_dimensions['D'].width = 14   # Date
    ws1.column_dimensions['E'].width = 16   # Status
    ws1.column_dimensions['G'].width = 14   # Summary Date
    ws1.column_dimensions['H'].width = 10   # Summary Count

    # Freeze header row
    ws1.freeze_panes = 'A2'

    wb.save(XLSX_TMP)
    print(f'Temporary XLSX created: {XLSX_TMP}')

    # Convert XLSX -> ODS using LibreOffice headless
    env = os.environ.copy()
    env['DISPLAY'] = ':0'
    result = subprocess.run(
        ['libreoffice', '--headless', '--convert-to', 'ods',
         '--outdir', WORKDIR, XLSX_TMP],
        capture_output=True, text=True, env=env, timeout=60
    )
    print('LibreOffice convert stdout:', result.stdout)
    print('LibreOffice convert stderr:', result.stderr)

    # The converted file will be named after the tmp xlsx base name
    converted = f'{WORKDIR}/{TASK_ID}_tmp.ods'
    if os.path.exists(converted):
        os.rename(converted, OUTPUT)
        print(f'Renamed {converted} -> {OUTPUT}')
    elif os.path.exists(OUTPUT):
        print(f'Output already at correct path: {OUTPUT}')
    else:
        # Fallback: try to find any newly created .ods
        import glob
        ods_files = glob.glob(f'{WORKDIR}/*.ods')
        if ods_files:
            newest = max(ods_files, key=os.path.getmtime)
            os.rename(newest, OUTPUT)
            print(f'Fallback rename {newest} -> {OUTPUT}')

    # Clean up temp xlsx
    if os.path.exists(XLSX_TMP):
        os.remove(XLSX_TMP)
        print(f'Removed temp file: {XLSX_TMP}')

    if os.path.exists(OUTPUT):
        print(f'Initial ODS file created: {OUTPUT}  ({os.path.getsize(OUTPUT)} bytes)')
    else:
        print(f'ERROR: Output file not found at {OUTPUT}')

    # ---------------------------------------------------------------
    # GUI-ready startup: open Chrome then LibreOffice Calc
    # ---------------------------------------------------------------
    # Kill any leftover soffice/chrome processes to ensure clean state
    subprocess.run(['pkill', '-f', 'soffice'], capture_output=True)
    time.sleep(1)

    # Launch Chrome (navigate to ArXiv cs.CL listings)
    launch_gui(
        'google-chrome --no-first-run --disable-session-crashed-bubble '
        '"https://arxiv.org/list/cs.CL/2024-01-15"',
        delay_sec=3.0
    )

    # Launch LibreOffice Calc with the tracker file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)

    print('GUI_READY: launched Chrome and LibreOffice Calc with DISPLAY=:0')


create_initial()
