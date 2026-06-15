"""
Initial Setup: Inventory export with Warehouse column having blanks for all rows
             except the first row of each warehouse group.
Task ID: osworld_calc_fill_blanks_above_004
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_fill_blanks_above_004'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Inventory ---
    ws = wb.active
    ws.title = 'Inventory'

    # Headers: Item SKU (A), Warehouse (B), Product Name (C), Stock (D)
    headers = ['Item SKU', 'Warehouse', 'Product Name', 'Stock']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Inventory data: Warehouse column only filled on first row of each group
    # Groups (NOT sorted alphabetically intentionally):
    #   Sunrise Depot (4 items), Central Hub (3 items),
    #   Metro Storage (4 items), Harbor Yard (3 items), Westside Cache (3 items)
    # Total 17 data rows

    data_full = [
        # Sunrise Depot group (rows 2-5)
        ('SKU-1041', 'Sunrise Depot', 'Industrial Drill Bit Set',  320),
        ('SKU-1042', 'Sunrise Depot', 'Safety Goggles Pro',        540),
        ('SKU-1043', 'Sunrise Depot', 'Heavy-Duty Work Gloves',    210),
        ('SKU-1044', 'Sunrise Depot', 'Adjustable Wrench 12in',    175),
        # Central Hub group (rows 6-8)
        ('SKU-2031', 'Central Hub',   'Laptop Stand Aluminum',     88),
        ('SKU-2032', 'Central Hub',   'USB-C Hub 7-Port',          260),
        ('SKU-2033', 'Central Hub',   'Wireless Keyboard Compact', 145),
        # Metro Storage group (rows 9-12)
        ('SKU-3011', 'Metro Storage', 'Garden Hose 50ft',          390),
        ('SKU-3012', 'Metro Storage', 'Stainless Steel Shovel',    115),
        ('SKU-3013', 'Metro Storage', 'Compost Bin 80L',           74),
        ('SKU-3014', 'Metro Storage', 'Pruning Shears Heavy Duty', 198),
        # Harbor Yard group (rows 13-15)
        ('SKU-4021', 'Harbor Yard',   'Marine Rope 20m',           430),
        ('SKU-4022', 'Harbor Yard',   'Waterproof Tarp 12x16ft',   220),
        ('SKU-4023', 'Harbor Yard',   'Anchor Chain 3m',           95),
        # Westside Cache group (rows 16-18)
        ('SKU-5001', 'Westside Cache','Camping Lantern LED',       310),
        ('SKU-5002', 'Westside Cache','Sleeping Bag -10C',         180),
        ('SKU-5003', 'Westside Cache','Portable Camp Stove',       142),
    ]

    # Write rows: Warehouse cell is blank except on the first item of each group
    current_warehouse = None
    for r, (sku, warehouse, product, stock) in enumerate(data_full, 2):
        ws.cell(row=r, column=1, value=sku)
        # Only write warehouse name on first row of each group
        if warehouse != current_warehouse:
            ws.cell(row=r, column=2, value=warehouse)
            current_warehouse = warehouse
        # else: leave column B blank (default None)
        ws.cell(row=r, column=3, value=product)
        ws.cell(row=r, column=4, value=stock)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 32
    ws.column_dimensions['D'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open LibreOffice Calc with the initial file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
