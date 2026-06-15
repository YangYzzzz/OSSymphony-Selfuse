"""
Reward Script: Fill blank Warehouse cells and sort by Warehouse alphabetically
Task ID: osworld_calc_fill_blanks_above_004
Domain: libreoffice_calc
Scoring:
  Component 1: All Warehouse cells in column B (rows 2-18) are filled (no blanks) — 0.5 points
  Component 2: Data rows (2-18) are sorted alphabetically by the Warehouse column — 0.5 points
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_fill_blanks_above_004'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook — fail fast if file cannot be opened
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Use the active sheet (should be 'Inventory')
    try:
        ws = wb.active
        print(f"INFO: Active sheet is '{ws.title}', dimensions: {ws.dimensions}")
    except Exception as e:
        print(f"CRITICAL: Cannot access active sheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: verify the sheet has data and the header row is correct
    try:
        header_a = ws.cell(row=1, column=1).value
        header_b = ws.cell(row=1, column=2).value
        max_row = ws.max_row
        if max_row < 2:
            print("CRITICAL: Sheet has no data rows. Aborting.")
            print("REWARD: 0.0")
            return 0.0
        print(f"INFO: Headers: A1={header_a}, B1={header_b}, data rows: {max_row - 1}")
    except Exception as e:
        print(f"CRITICAL: Cannot read header row: {e}")
        print("REWARD: 0.0")
        return 0.0

    # -------------------------------------------------------------------------
    # Component 1: All Warehouse cells in column B (rows 2 to max_row) are filled
    # (i.e., no None/blank values). This verifies the "fill down" action.
    # Expected: 0.5 points earned on golden, 0 on initial (initial has blanks).
    # -------------------------------------------------------------------------
    try:
        blank_cells = []
        warehouse_values = []
        for row_idx in range(2, max_row + 1):
            cell_val = ws.cell(row=row_idx, column=2).value
            warehouse_values.append((row_idx, cell_val))
            if cell_val is None or str(cell_val).strip() == '':
                blank_cells.append(row_idx)

        if len(blank_cells) == 0:
            print(f"PASS: Component 1 — All Warehouse cells are filled (no blanks in B2:B{max_row}) (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — Found {len(blank_cells)} blank Warehouse cells at rows: {blank_cells[:10]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        warehouse_values = []

    # -------------------------------------------------------------------------
    # Component 2: Data is sorted alphabetically by the Warehouse column (B).
    # We extract all warehouse values from rows 2..max_row and verify they are
    # in non-decreasing alphabetical order (case-insensitive).
    # Expected: 0.5 points earned on golden, 0 on initial (initial is not sorted).
    # -------------------------------------------------------------------------
    try:
        wh_names = [str(val).strip() if val is not None else '' for _, val in warehouse_values]

        # Check if the list is sorted alphabetically (case-insensitive)
        is_sorted = all(
            wh_names[i].lower() <= wh_names[i + 1].lower()
            for i in range(len(wh_names) - 1)
        )

        if is_sorted and len(wh_names) > 0:
            print(f"PASS: Component 2 — Data is sorted alphabetically by Warehouse "
                  f"(first: '{wh_names[0]}', last: '{wh_names[-1]}') (0.5 pts)")
            total_score += 0.5
        else:
            # Find the first out-of-order pair for diagnosis
            first_disorder = None
            for i in range(len(wh_names) - 1):
                if wh_names[i].lower() > wh_names[i + 1].lower():
                    first_disorder = (i + 2, wh_names[i], i + 3, wh_names[i + 1])
                    break
            if first_disorder:
                print(f"FAIL: Component 2 — Data not sorted. "
                      f"Row {first_disorder[0]} has '{first_disorder[1]}' "
                      f"but Row {first_disorder[2]} has '{first_disorder[3]}'")
            else:
                print(f"FAIL: Component 2 — Sorting check failed (empty data or unexpected state)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the given env
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
