"""
Reward Script: Replace lowercase 'new' with 'updated' in product descriptions
Task ID: calc_dop_findreplace_matchcase_030
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): All 12 lowercase 'new' occurrences in col C are replaced (count == 0)
  Component 2 (0.5): Exactly 12 'updated' occurrences appear in col C (correct replacements)

NOTE: Components 2 and 3 from an earlier design checked that 'New' and 'NEW'
were preserved — but those conditions also hold in the initial file (they never
changed), making them pre-existing properties that MUST NOT contribute to the score.
Only changes between initial and golden are scored here.
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_dop_findreplace_matchcase_030'
SHEET_NAME = 'Products'
DATA_START_ROW = 2
DATA_END_ROW = 61
DESC_COL = 3  # Column C


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Only task-introduced changes are scored:
    - Initial file: 12 lowercase 'new' in col C, 0 'updated'
    - Golden file:  0 lowercase 'new' in col C, 12 'updated'
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: sheet 'Products' must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Collect all description values from column C (rows 2-61)
    descriptions = []
    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        val = ws.cell(row=row, column=DESC_COL).value
        descriptions.append((row, str(val) if val is not None else ''))

    # Component 1: Zero lowercase 'new' occurrences remain in col C (0.5 points)
    # In the initial file there are 12 lowercase 'new'; task requires all to be replaced.
    # This FAILS on initial (12 found) and PASSES on golden (0 found).
    try:
        lowercase_new_count = 0
        rows_with_lowercase_new = []

        for row, text in descriptions:
            # Match standalone 'new' where all three letters are lowercase
            # re.findall with \bnew\b only matches 'new', not 'New' or 'NEW'
            matches = re.findall(r'\bnew\b', text)
            if matches:
                lowercase_new_count += len(matches)
                rows_with_lowercase_new.append(row)

        if lowercase_new_count == 0:
            print(f"PASS: Component 1 — No lowercase 'new' occurrences remain in column C. (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — Found {lowercase_new_count} lowercase 'new' occurrence(s) "
                  f"still present (expected 0). Rows: {rows_with_lowercase_new}.")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Exactly 12 'updated' occurrences appear in col C (0.5 points)
    # In the initial file there are 0 'updated'; task requires exactly 12 replacements.
    # This FAILS on initial (0 found) and PASSES on golden (12 found).
    try:
        updated_count = 0
        rows_with_updated = []

        for row, text in descriptions:
            # Match standalone 'updated' (exact case as expected replacement)
            matches = re.findall(r'\bupdated\b', text)
            if matches:
                updated_count += len(matches)
                rows_with_updated.append(row)

        if updated_count == 12:
            print(f"PASS: Component 2 — Exactly 12 'updated' occurrences found in column C "
                  f"at rows {rows_with_updated}. (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 2 — Expected 12 'updated' occurrences in column C, "
                  f"found {updated_count}. Rows: {rows_with_updated}.")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
