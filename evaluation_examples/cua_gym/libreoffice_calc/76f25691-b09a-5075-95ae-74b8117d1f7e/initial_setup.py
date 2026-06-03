"""
Initial Setup: Find & Replace with Match Case — lowercase 'new' only
Task ID: calc_dop_findreplace_matchcase_030
Domain: libreoffice_calc
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_dop_findreplace_matchcase_030'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Products ---
    ws = wb.active
    ws.title = 'Products'

    # Headers: Product ID (A), Name (B), Description (C), Tags (D)
    headers = ['Product ID', 'Name', 'Description', 'Tags']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Product data with realistic content.
    # Column C descriptions:
    #   - 12 rows with lowercase 'new'       (should be replaced → 'updated')
    #   - 5  rows with title-case 'New'      (should NOT be replaced)
    #   - 3  rows with all-caps 'NEW'        (should NOT be replaced)
    #   - remaining rows have no 'new' variant
    #
    # Total rows: 60 (rows 2–61)
    data = [
        # --- 12 lowercase 'new' occurrences (rows 2–13) ---
        ('PRD-001', 'Wireless Keyboard',        'Featuring new ergonomic layout for comfort',         'electronics,input'),
        ('PRD-002', 'Standing Desk',            'new height-adjustable frame with memory settings',   'furniture,office'),
        ('PRD-003', 'USB-C Hub 7-Port',         'Compact new design fits any laptop model',           'electronics,hub'),
        ('PRD-004', 'Noise-Cancelling Headset', 'Includes new active noise reduction technology',     'audio,wireless'),
        ('PRD-005', 'Smart Thermostat Pro',     'Built on a new sensor platform for accuracy',        'smart-home,energy'),
        ('PRD-006', 'LED Desk Lamp',            'Energy-saving new bulb included in package',         'lighting,office'),
        ('PRD-007', 'Portable SSD 1TB',         'new faster NVMe interface for rapid transfers',      'storage,portable'),
        ('PRD-008', 'Bluetooth Speaker',        'Upgraded with new 360-degree surround sound',        'audio,portable'),
        ('PRD-009', 'Mechanical Keyboard',      'Ships with new tactile brown switches installed',    'electronics,input'),
        ('PRD-010', 'Gaming Mouse',             'Fitted with new optical sensor at 16000 DPI',        'gaming,input'),
        ('PRD-011', 'Webcam 4K',                'Uses new low-light autofocus algorithm',             'video,conference'),
        ('PRD-012', 'Laptop Backpack',          'Redesigned with new waterproof external fabric',     'accessories,bag'),

        # --- 5 title-case 'New' occurrences (rows 14–18) ---
        ('PRD-013', 'Monitor 27-inch',          'New IPS panel with 165 Hz refresh rate',             'display,gaming'),
        ('PRD-014', 'Ergonomic Chair',          'New lumbar support design reduces back strain',       'furniture,health'),
        ('PRD-015', 'Wireless Mouse',           'New scroll wheel with magnetic precision glide',      'input,wireless'),
        ('PRD-016', 'Smart Plug Duo',           'New scheduling firmware pre-installed',               'smart-home,energy'),
        ('PRD-017', 'Drawing Tablet A4',        'New pressure-sensitive stylus included',              'creative,input'),

        # --- 3 all-caps 'NEW' occurrences (rows 19–21) ---
        ('PRD-018', 'NVMe SSD 2TB',             'NEW ultra-speed controller for professionals',        'storage,internal'),
        ('PRD-019', 'OLED Monitor 34-inch',     'Flagship display — NEW for 2025 lineup',              'display,premium'),
        ('PRD-020', 'Gaming Headset 7.1',       'NEW surround audio engine with haptic bass',          'audio,gaming'),

        # --- 39 rows with no 'new' variant (rows 22–60) ---
        ('PRD-021', 'Cable Management Kit',     'Organise cables with hook-and-loop velcro ties',     'accessories,cables'),
        ('PRD-022', 'Monitor Arm Single',       'Full articulating arm supports up to 32 inches',     'display,ergonomics'),
        ('PRD-023', 'USB-A 3.0 Hub 4-Port',    'Plug-and-play hub for legacy USB-A devices',         'electronics,hub'),
        ('PRD-024', 'Desk Pad XL',             'Non-slip base with smooth cloth surface 90x45cm',    'accessories,desk'),
        ('PRD-025', 'Surge Protector 6-Outlet','6 outlets with 2400-joule surge suppression',        'power,safety'),
        ('PRD-026', 'Streaming Microphone',    'Cardioid condenser mic for podcasting and gaming',   'audio,recording'),
        ('PRD-027', 'Graphic Tablet Pro',      'Pressure-sensitive stylus with 8192 levels',         'creative,input'),
        ('PRD-028', 'Keyboard Wrist Rest',     'Memory foam pad reduces repetitive-strain risk',     'accessories,ergonomics'),
        ('PRD-029', 'Conference Speakerphone', '360-degree mic array for huddle-room meetings',      'audio,conference'),
        ('PRD-030', 'Privacy Screen Filter',   'Blackout side-angle privacy protection 15.6 inch',  'accessories,security'),
        ('PRD-031', 'External DVD Drive',      'USB-powered slim drive for burning and reading',     'storage,optical'),
        ('PRD-032', 'Cable HDMI 2.1 3m',       'Supports 4K 120 Hz and 8K 60 Hz signal',            'cables,display'),
        ('PRD-033', 'KVM Switch 2-Port',       'Switch keyboard/video/mouse between two computers', 'networking,input'),
        ('PRD-034', 'HD Webcam 1080p',         'Built-in ring light for flattering video calls',    'video,conference'),
        ('PRD-035', 'Laptop Cooling Pad',      'Dual fan cooling for 15-17 inch laptops',           'accessories,cooling'),
        ('PRD-036', 'VESA Mount Bracket',      'Universal 75mm and 100mm VESA compatibility',       'display,mounting'),
        ('PRD-037', 'Portable Projector',      'HD resolution with 120-inch projected image',       'display,presentation'),
        ('PRD-038', 'Label Printer',           'Thermal label printing at 203 DPI resolution',      'office,printing'),
        ('PRD-039', 'Desk Clock Digital',      'Large display with room temperature readout',       'office,accessories'),
        ('PRD-040', 'Wireless Charging Pad',   'Qi-certified fast charging pad 15W',                'accessories,charging'),
        ('PRD-041', 'Smart Power Strip',       'App-controlled individual outlet scheduling',       'smart-home,power'),
        ('PRD-042', 'HDMI Capture Card',       'Record gameplay at 1080p 60fps via USB 3.0',       'gaming,video'),
        ('PRD-043', 'Touchscreen Monitor 24in','Multi-touch 10-point IPS display for kiosks',      'display,interactive'),
        ('PRD-044', 'Rollable Cable Mat',      'Protect cables and provide tidy workspace',         'accessories,cables'),
        ('PRD-045', 'LED Strip Lights 5m',     'App-controlled RGB lighting for behind monitors',  'lighting,smart-home'),
        ('PRD-046', 'Laptop Lock Cable',       'Security cable with combination lock 1.5m',        'accessories,security'),
        ('PRD-047', 'USB-C Power Bank 26800mAh','Fast charging power bank with dual USB-C output','accessories,charging'),
        ('PRD-048', 'Fingerprint Reader USB',  'One-touch login for Windows Hello biometrics',     'security,input'),
        ('PRD-049', 'Portable Monitor 15.6in', 'Full-HD USB-C secondary display weighs 780g',      'display,portable'),
        ('PRD-050', 'Smart Badge Printer',     'Print ID cards and event badges directly',         'office,printing'),
        ('PRD-051', 'Foot Rest Ergonomic',     'Adjustable tilt and height foot support',          'furniture,ergonomics'),
        ('PRD-052', 'Conference Camera 180°',  'Ultra-wide lens covers full meeting room',         'video,conference'),
        ('PRD-053', 'Desk Organiser Bamboo',   'Eco-friendly bamboo compartment desk organiser',   'office,accessories'),
        ('PRD-054', 'Screen Cleaner Kit',      'Microfibre cloth and streak-free solution',        'accessories,cleaning'),
        ('PRD-055', 'Dual Monitor Stand',      'Freestanding grommet mount for two monitors',      'display,ergonomics'),
        ('PRD-056', 'Barcode Scanner USB',     'Omnidirectional scan for inventory management',    'office,scanning'),
        ('PRD-057', 'Presentation Remote',     'Laser pointer remote with 100m wireless range',   'office,presentation'),
        ('PRD-058', 'Laptop Docking Station',  'Thunderbolt 4 dock with 96W charging passthrough','electronics,hub'),
        ('PRD-059', 'Smart Whiteboard 65in',   'Interactive display with cloud sync capabilities','office,display'),
        ('PRD-060', 'Paper Shredder Micro-Cut','6-sheet micro-cut shredder with safety lock',     'office,security'),
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        ws.cell(row=r, column=2, value=row_data[1])
        ws.cell(row=r, column=3, value=row_data[2])
        ws.cell(row=r, column=4, value=row_data[3])

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Quick verification: count 'new' variants in column C
    wb2 = openpyxl.load_workbook(OUTPUT)
    ws2 = wb2['Products']
    lowercase_count = 0
    titlecase_count = 0
    uppercase_count = 0
    for row in ws2.iter_rows(min_row=2, max_row=61, min_col=3, max_col=3):
        val = row[0].value or ''
        import re
        # count standalone 'new' (lowercase) - word boundary check
        lowercase_count += len(re.findall(r'\bnew\b', val))
        titlecase_count += len(re.findall(r'\bNew\b', val))
        uppercase_count += len(re.findall(r'\bNEW\b', val))
    print(f"Lowercase 'new' occurrences: {lowercase_count} (expected 12)")
    print(f"Title-case 'New' occurrences: {titlecase_count} (expected 5)")
    print(f"All-caps 'NEW' occurrences: {uppercase_count} (expected 3)")

create_initial()
