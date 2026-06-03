"""
Initial Setup: Logistics cost breakdown with pivot-style SUMIFS calculations
Task ID: calc_ops_050
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_050'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Data ---
    ws_data = wb.active
    ws_data.title = 'Data'

    headers = ['Date', 'Carrier', 'Service', 'Cost']
    for col, h in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    data = [
        [date(2026, 1, 5),  'FedEx',  'Express', 450],
        [date(2026, 1, 12), 'UPS',    'Ground',  180],
        [date(2026, 1, 20), 'FedEx',  'Ground',  220],
        [date(2026, 2, 3),  'UPS',    'Express', 380],
        [date(2026, 2, 15), 'FedEx',  'Express', 510],
        [date(2026, 2, 22), 'DHL',    'Express', 620],
        [date(2026, 3, 5),  'FedEx',  'Ground',  195],
        [date(2026, 3, 18), 'UPS',    'Ground',  210],
        [date(2026, 3, 25), 'DHL',    'Express', 575],
    ]

    for r, row_data in enumerate(data, 2):
        ws_data.cell(row=r, column=1, value=row_data[0])
        ws_data.cell(row=r, column=1).number_format = 'yyyy-mm-dd'
        ws_data.cell(row=r, column=2, value=row_data[1])
        ws_data.cell(row=r, column=3, value=row_data[2])
        ws_data.cell(row=r, column=4, value=row_data[3])

    # Set column widths for readability
    ws_data.column_dimensions['A'].width = 14
    ws_data.column_dimensions['B'].width = 12
    ws_data.column_dimensions['C'].width = 12
    ws_data.column_dimensions['D'].width = 10

    # --- Sheet 2: Summary ---
    ws_summary = wb.create_sheet('Summary')

    summary_headers = ['Carrier', 'Jan', 'Feb', 'Mar', 'Total']
    for col, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    carriers = ['FedEx', 'UPS', 'DHL']
    for r, carrier in enumerate(carriers, 2):
        ws_summary.cell(row=r, column=1, value=carrier)

    # Summary B2:E4 left EMPTY - that's the task for the agent to fill with SUMIFS formulas

    ws_summary.column_dimensions['A'].width = 12
    ws_summary.column_dimensions['B'].width = 10
    ws_summary.column_dimensions['C'].width = 10
    ws_summary.column_dimensions['D'].width = 10
    ws_summary.column_dimensions['E'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
