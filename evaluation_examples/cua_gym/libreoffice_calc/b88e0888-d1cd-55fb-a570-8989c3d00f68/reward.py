"""
Reward Script: Logistics cost breakdown analysis with pivot-style SUMIFS calculations
Task ID: calc_ops_050
Domain: libreoffice_calc
Scoring:
  Component 1 (0.50): Monthly cost formulas in B2:D4 — each cell contains a formula
                       referencing carrier and month to aggregate costs from Data sheet
  Component 2 (0.20): Total formulas in E2:E4 — each cell contains a SUM formula
  Component 3 (0.30): Formula correctness — formulas produce correct ground-truth values
                       (verified via LibreOffice save + data_only read)
"""

import os
import re
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_050'

# Ground truth from task context
EXPECTED_VALUES = {
    # (row, col): expected_value
    # Row 2 = FedEx, Row 3 = UPS, Row 4 = DHL
    # Col B=Jan, C=Feb, D=Mar, E=Total
    ('B2',): 670,  ('C2',): 510,  ('D2',): 195,  ('E2',): 1375,
    ('B3',): 180,  ('C3',): 380,  ('D3',): 210,  ('E3',): 770,
    ('B4',): 0,    ('C4',): 620,  ('D4',): 575,  ('E4',): 1195,
}

# Map month columns to expected month numbers
MONTH_MAP = {'B': 1, 'C': 2, 'D': 3}

# Carrier rows
CARRIER_MAP = {2: 'FedEx', 3: 'UPS', 4: 'DHL'}


def persist_app_state():
    """Save any unsaved LibreOffice edits via Ctrl+S."""
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(1.0)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def is_monthly_cost_formula(formula_str, row_num, col_letter):
    """
    Check if the formula aggregates costs by carrier and month.
    Accepts SUMIFS, SUMPRODUCT, or equivalent patterns that reference:
    - The carrier column (B) matching the carrier in column A of this row
    - The date column (A) filtered by month
    - The cost column (D)
    """
    if not isinstance(formula_str, str) or not formula_str.startswith('='):
        return False

    f = formula_str.upper()
    month_num = MONTH_MAP.get(col_letter)

    # Must reference cost data (column D from Data sheet)
    has_cost_ref = bool(re.search(r'D\$?\d', f) or re.search(r'DATA[.!]D', f, re.IGNORECASE))

    # Must reference carrier column (B from Data sheet) or $A for the carrier name
    has_carrier_ref = bool(re.search(r'B\$?\d', f) or re.search(r'DATA[.!]B', f, re.IGNORECASE))

    # Must reference date/month — either MONTH() function or month number
    has_month_ref = bool(
        re.search(r'MONTH\s*\(', f) or
        (month_num and str(month_num) in f)
    )

    # Must be an aggregation formula (SUMIFS, SUMPRODUCT, or similar)
    has_agg_func = bool(re.search(r'(SUMIFS?|SUMPRODUCT|SUMIF)\s*\(', f))

    return has_cost_ref and has_carrier_ref and has_month_ref and has_agg_func


def is_total_formula(formula_str, row_num):
    """
    Check if the formula sums the monthly values for this row.
    Accepts =SUM(B<row>:D<row>) or equivalent.
    """
    if not isinstance(formula_str, str) or not formula_str.startswith('='):
        return False

    f = formula_str.upper().replace(' ', '')
    # Accept SUM of the row's monthly cells
    if re.search(r'SUM\s*\(', f):
        # Check it references B, C, or D of the same row
        if re.search(rf'B{row_num}', f) and re.search(rf'D{row_num}', f):
            return True
        # Also accept range like B2:D2
        if re.search(rf'B{row_num}\s*:\s*D{row_num}', f):
            return True
    # Accept simple addition: =B2+C2+D2
    if re.search(rf'B{row_num}\s*\+\s*C{row_num}\s*\+\s*D{row_num}', f):
        return True

    return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Summary sheet must exist
    if 'Summary' not in wb.sheetnames:
        print("CRITICAL: 'Summary' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Summary']

    # =========================================================================
    # Component 1: Monthly cost formulas in B2:D4 (0.50 points)
    # Each of the 9 cells should contain a formula aggregating costs by
    # carrier and month. This is the core task requirement.
    # =========================================================================
    try:
        monthly_pass = 0
        monthly_total = 9  # 3 carriers x 3 months
        for row_num in [2, 3, 4]:
            for col_letter in ['B', 'C', 'D']:
                coord = f'{col_letter}{row_num}'
                val = ws[coord].value
                if is_monthly_cost_formula(val, row_num, col_letter):
                    monthly_pass += 1
                    print(f"  PASS: {coord} has valid monthly formula: {val}")
                else:
                    print(f"  FAIL: {coord} — expected monthly aggregation formula, found: {val!r}")

        if monthly_pass == monthly_total:
            print(f"PASS: Component 1 — All {monthly_total} monthly formulas present (0.50 pts)")
            total_score += 0.50
        elif monthly_pass > 0:
            partial = round(0.50 * (monthly_pass / monthly_total), 2)
            print(f"PARTIAL: Component 1 — {monthly_pass}/{monthly_total} monthly formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No valid monthly formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Total formulas in E2:E4 (0.20 points)
    # Each carrier row should have a SUM formula in column E.
    # =========================================================================
    try:
        total_pass = 0
        total_cells = 3
        for row_num in [2, 3, 4]:
            coord = f'E{row_num}'
            val = ws[coord].value
            if is_total_formula(val, row_num):
                total_pass += 1
                print(f"  PASS: {coord} has valid total formula: {val}")
            else:
                print(f"  FAIL: {coord} — expected SUM formula, found: {val!r}")

        if total_pass == total_cells:
            print(f"PASS: Component 2 — All {total_cells} total formulas present (0.20 pts)")
            total_score += 0.20
        elif total_pass > 0:
            partial = round(0.20 * (total_pass / total_cells), 2)
            print(f"PARTIAL: Component 2 — {total_pass}/{total_cells} total formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No valid total formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Formula correctness via computed values (0.30 points)
    # Use LibreOffice to compute formula values, then verify against ground truth.
    # If LibreOffice hasn't computed values (data_only returns None), fall back
    # to verifying formula reference ranges cover data rows 2-10.
    # =========================================================================
    try:
        # Try loading with data_only to get cached computed values
        wb_data = openpyxl.load_workbook(file_path, data_only=True)
        ws_data = wb_data['Summary']

        # Check if any computed values are available
        sample_val = ws_data['B2'].value
        has_computed = sample_val is not None

        if has_computed:
            # Verify computed values against ground truth
            correct = 0
            total_checks = 12  # 3 carriers x 4 columns
            for coord_tuple, expected in EXPECTED_VALUES.items():
                coord = coord_tuple[0]
                actual = ws_data[coord].value
                if actual is not None:
                    try:
                        if abs(float(actual) - expected) < 0.01:
                            correct += 1
                            print(f"  PASS: {coord} = {actual} (expected {expected})")
                        else:
                            print(f"  FAIL: {coord} = {actual} (expected {expected})")
                    except (ValueError, TypeError):
                        print(f"  FAIL: {coord} = {actual!r} (not numeric, expected {expected})")
                else:
                    print(f"  FAIL: {coord} = None (expected {expected})")

            if correct == total_checks:
                print(f"PASS: Component 3 — All {total_checks} values correct (0.30 pts)")
                total_score += 0.30
            elif correct > 0:
                partial = round(0.30 * (correct / total_checks), 2)
                print(f"PARTIAL: Component 3 — {correct}/{total_checks} values correct ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 3 — No values matched ground truth")
        else:
            print("INFO: No cached computed values available (file not opened in LibreOffice)")
            print("INFO: Falling back to formula range verification")

            # Fallback: verify formulas reference the correct data range (rows 2-10)
            range_correct = 0
            range_total = 9  # monthly formulas only
            for row_num in [2, 3, 4]:
                for col_letter in ['B', 'C', 'D']:
                    coord = f'{col_letter}{row_num}'
                    val = ws[coord].value
                    if isinstance(val, str):
                        f_upper = val.upper()
                        # Check that the formula references the full data range (at least row 2 to row 10)
                        if re.search(r'2.*10|2:.*10|\$2.*\$10', f_upper):
                            range_correct += 1
                            print(f"  PASS: {coord} formula covers data range 2-10")
                        else:
                            print(f"  FAIL: {coord} formula may not cover full data range: {val}")
                    else:
                        print(f"  FAIL: {coord} is not a formula: {val!r}")

            if range_correct == range_total:
                print(f"PASS: Component 3 (fallback) — All formulas reference correct range (0.30 pts)")
                total_score += 0.30
            elif range_correct > 0:
                partial = round(0.30 * (range_correct / range_total), 2)
                print(f"PARTIAL: Component 3 (fallback) — {range_correct}/{range_total} correct ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 3 (fallback) — No formulas reference correct data range")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
