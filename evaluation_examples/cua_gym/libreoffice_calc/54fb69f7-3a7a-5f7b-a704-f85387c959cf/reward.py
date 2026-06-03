"""
Reward Script: Fix VLOOKUP type mismatch (#N/A due to numeric vs text IDs)
Task ID: calc_tbl_009
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): C2 formula wraps lookup value with TEXT() or equivalent type conversion
  Component 2 (0.3): All other VLOOKUP formulas (C3:C13) also fixed with type conversion
  Component 3 (0.2): All fixed formulas still correctly reference Sheet2 lookup range
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_009'


def persist_app_state(domain: str):
    """Best-effort save of any unsaved GUI edits."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def formula_has_type_conversion(formula):
    """
    Check if a VLOOKUP formula handles the numeric-to-text type mismatch.
    Accepted patterns include:
      - TEXT(A2,"0") or TEXT(A2,"#") etc.  (convert number to text)
      - A2&""  (concatenate empty string to coerce to text)
      - ""&A2
      - TRIM(A2) patterns that coerce type
      - Any wrapping that converts the first argument from raw cell ref to text
    The key requirement: the first argument to VLOOKUP must NOT be a bare cell ref
    like A2, A3, etc. It must be wrapped in some conversion function or expression.
    """
    if not formula or not isinstance(formula, str):
        return False

    upper = formula.upper().replace(" ", "")

    # Must still be a VLOOKUP
    if "VLOOKUP(" not in upper:
        return False

    # Extract the first argument of VLOOKUP (everything between VLOOKUP( and the first comma
    # that is not inside nested parentheses)
    match = re.search(r'VLOOKUP\(', upper)
    if not match:
        return False

    start = match.end()
    depth = 1
    pos = start
    while pos < len(upper) and depth > 0:
        ch = upper[pos]
        if ch == '(':
            depth += 1
        elif ch == ')':
            depth -= 1
        elif ch == ',' and depth == 1:
            break
        pos += 1

    first_arg = upper[start:pos]

    # The original broken formula has a bare cell ref like "A2" as first arg.
    # The fix should wrap it in TEXT(), concatenate with "", or use some other conversion.
    # Check that the first arg is NOT just a bare cell reference (e.g. A2, A3, ..., A13)
    if re.fullmatch(r'[A-Z]+\d+', first_arg):
        # Bare cell reference — NOT fixed
        return False

    # If first arg contains TEXT( or string concatenation (&"") or VALUE( or similar, it's fixed
    # We already confirmed it's not bare, so any wrapping counts as a fix attempt
    return True


def verify_task(file_path):
    """
    Verify that the VLOOKUP formulas in column C handle the numeric-to-text type mismatch.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify 'Orders' sheet exists
    if 'Orders' not in wb.sheetnames:
        print("CRITICAL: 'Orders' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Orders']

    # Component 1: C2 formula uses type conversion (0.5 points)
    # This is the primary cell mentioned in the task instruction.
    try:
        c2_val = ws['C2'].value
        print(f"DEBUG: C2 value = {repr(c2_val)}")
        if formula_has_type_conversion(c2_val):
            print(f"PASS: Component 1 — C2 formula has type conversion: {c2_val} (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — C2 formula lacks type conversion: {c2_val}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All other VLOOKUP formulas C3:C13 also fixed (0.3 points)
    # The task says "fix the formula" but all C column formulas have the same issue.
    # A thorough fix should address all of them, not just C2.
    try:
        fixed_count = 0
        total_formulas = 0
        for row in range(3, 14):  # C3 through C13
            cell_val = ws.cell(row=row, column=3).value
            if cell_val and isinstance(cell_val, str) and 'VLOOKUP' in cell_val.upper():
                total_formulas += 1
                if formula_has_type_conversion(cell_val):
                    fixed_count += 1
                else:
                    print(f"  DETAIL: C{row} not fixed: {cell_val}")

        print(f"DEBUG: {fixed_count}/{total_formulas} additional formulas fixed (C3:C13)")

        if total_formulas > 0 and fixed_count == total_formulas:
            print(f"PASS: Component 2 — All {total_formulas} additional VLOOKUP formulas fixed (0.3 pts)")
            total_score += 0.3
        elif total_formulas > 0 and fixed_count > 0:
            partial = 0.3 * (fixed_count / total_formulas)
            print(f"PARTIAL: Component 2 — {fixed_count}/{total_formulas} fixed ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No additional formulas fixed")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Fixed formulas still reference Sheet2 correctly AND have type conversion (0.2 points)
    # Anchored to the task change: only counts formulas that BOTH have type conversion
    # AND retain the correct Sheet2 lookup reference. This fails on initial_env because
    # no formulas have type conversion there.
    try:
        valid_count = 0
        checked = 0
        for row in range(2, 14):  # C2 through C13
            cell_val = ws.cell(row=row, column=3).value
            if cell_val and isinstance(cell_val, str):
                upper_val = cell_val.upper().replace(" ", "")
                checked += 1
                has_conversion = formula_has_type_conversion(cell_val)
                has_sheet2_ref = 'SHEET2' in upper_val and ('A:B' in upper_val or 'A1:B' in upper_val)
                if has_conversion and has_sheet2_ref:
                    valid_count += 1
                elif has_conversion and not has_sheet2_ref:
                    print(f"  DETAIL: C{row} has type conversion but broken Sheet2 ref: {cell_val}")

        print(f"DEBUG: {valid_count}/{checked} formulas have BOTH type conversion AND Sheet2 ref")

        if checked > 0 and valid_count == checked:
            print(f"PASS: Component 3 — All fixed formulas correctly reference Sheet2.A:B (0.2 pts)")
            total_score += 0.2
        elif checked > 0 and valid_count > 0:
            partial = 0.2 * (valid_count / checked)
            print(f"PARTIAL: Component 3 — {valid_count}/{checked} valid ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No formulas have both type conversion and correct Sheet2 ref")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook — save any unsaved GUI state before verification
persist_app_state("libreoffice_calc")

# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
