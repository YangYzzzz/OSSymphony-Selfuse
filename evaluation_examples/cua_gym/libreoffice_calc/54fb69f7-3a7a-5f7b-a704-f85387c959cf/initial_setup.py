"""
Initial Setup: Fix VLOOKUP #N/A due to numeric/text type mismatch
Task ID: calc_tbl_009
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_009'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: Orders ---
    ws1 = wb.active
    ws1.title = 'Orders'

    headers1 = ['OrderID', 'OrderDate', 'CustomerName', 'Product', 'Quantity', 'UnitPrice', 'Total']
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)

    # OrderID values are NUMBERS (this is the key mismatch)
    orders = [
        [1001, '2025-03-01', None, 'Wireless Mouse', 3, 24.99, 74.97],
        [1002, '2025-03-02', None, 'USB-C Hub', 1, 45.50, 45.50],
        [1003, '2025-03-03', None, 'Mechanical Keyboard', 2, 89.00, 178.00],
        [1004, '2025-03-05', None, 'Monitor Stand', 1, 35.00, 35.00],
        [1005, '2025-03-06', None, '27" Monitor', 1, 329.99, 329.99],
        [1006, '2025-03-07', None, 'Webcam HD', 2, 59.99, 119.98],
        [1007, '2025-03-08', None, 'Desk Lamp LED', 4, 22.00, 88.00],
        [1008, '2025-03-10', None, 'Laptop Sleeve', 5, 18.50, 92.50],
        [1009, '2025-03-11', None, 'HDMI Cable 2m', 10, 8.99, 89.90],
        [1010, '2025-03-12', None, 'Noise-Cancel Headphones', 1, 199.00, 199.00],
        [1011, '2025-03-14', None, 'Ergonomic Chair Mat', 2, 42.00, 84.00],
        [1012, '2025-03-15', None, 'USB Flash Drive 64GB', 8, 12.99, 103.92],
    ]

    for r, row_data in enumerate(orders, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # C2:C13 — VLOOKUP that will fail due to type mismatch (numeric vs text)
    # The formula uses A2 (number 1001) to look up in Sheet2 A:B where A has TEXT IDs
    for r in range(2, 14):
        ws1.cell(row=r, column=3, value='=VLOOKUP(A{row},Sheet2.A:B,2,0)'.format(row=r))

    # --- Sheet2: Customers ---
    ws2 = wb.create_sheet('Sheet2')
    ws2['A1'] = 'CustomerID'
    ws2['B1'] = 'CustomerName'

    # Customer IDs stored as TEXT strings — this causes the #N/A mismatch
    customers = [
        ['1001', 'Sarah Chen'],
        ['1002', 'Marcus Johnson'],
        ['1003', 'Emily Rodriguez'],
        ['1004', 'James Okonkwo'],
        ['1005', 'Priya Sharma'],
        ['1006', 'David Kim'],
        ['1007', 'Olivia Martínez'],
        ['1008', 'Liam O\'Brien'],
        ['1009', 'Aisha Patel'],
        ['1010', 'Thomas Weber'],
        ['1011', 'Isabella Rossi'],
        ['1012', 'Nathan Brooks'],
    ]

    for r, (cid, cname) in enumerate(customers, 2):
        # Store CustomerID explicitly as text string
        ws2.cell(row=r, column=1, value=cid)
        ws2.cell(row=r, column=2, value=cname)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
