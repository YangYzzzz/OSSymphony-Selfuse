"""
Initial Setup: Sales Pricing Margin Guard
Task ID: calc_sales_pricing_margin_guard_036
Domain: libreoffice_calc

Creates the DealPricer spreadsheet with 40 deals.
Columns F-I (Net Price, Gross Profit, Margin %, Approval Status) are left empty.
No conditional formatting or comments applied.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_pricing_margin_guard_036'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'DealPricer'

    # --- Headers ---
    headers = [
        'Deal ID', 'Product', 'COGS', 'List Price',
        'Discount %', 'Net Price', 'Gross Profit', 'Margin %', 'Approval Status'
    ]
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    white_bold_font = Font(name='Calibri', bold=True, size=11, color='FFFFFFFF')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = white_bold_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[1].height = 22

    # --- Column widths ---
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 13
    ws.column_dimensions['F'].width = 13
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 26

    # --- Deal data: (product, COGS, List Price, Discount %) ---
    # Some deals have high COGS/LP ratios (~50-58%) with high discounts
    # so that margin drops below 10% (critically low) and below 20%
    deals = [
        ('Industrial Laser Cutter',       380,  950,  0.08),   # margin ~57%
        ('Portable Air Compressor',        90,  270,  0.12),   # margin ~63%
        ('Compact Camera System',          55,  100,  0.40),   # margin ~8%  CRITICAL
        ('Smart Thermostat Pro',           40,  130,  0.15),   # margin ~64%
        ('Wireless Barcode Scanner',       55,  160,  0.10),   # margin ~62%
        ('Electric Pallet Jack',          420,  900,  0.20),   # margin ~42%
        ('Server Rack Cabinet 12U',       280,  680,  0.25),   # margin ~45%
        ('Bulk Storage Drive 8TB',         66,  120,  0.42),   # margin ~5%  CRITICAL
        ('Digital Multimeter Pro',         28,   76,  0.12),   # margin ~58%
        ('Bench Drill Press 1/2in',       140,  360,  0.14),   # margin ~56%
        ('Thermal Label Printer',          60,  170,  0.18),   # margin ~57%
        ('UPS Battery Backup 1500VA',      92,  235,  0.16),   # margin ~53%
        ('Thermal Imaging Camera',        275,  500,  0.43),   # margin ~4%  CRITICAL
        ('Pneumatic Nail Gun 21-deg',      65,  175,  0.08),   # margin ~60%
        ('Commercial Vacuum 6gal',        118,  290,  0.09),   # margin ~55%
        ('Network Switch 48-Port',        165,  380,  0.22),   # margin ~44%
        ('Oscilloscope 100MHz',           165,  300,  0.41),   # margin ~7%  CRITICAL
        ('Electric Chain Hoist 1T',       195,  460,  0.10),   # margin ~53%
        ('Fiber Patch Panel 24-Port',      72,  195,  0.08),   # margin ~62%
        ('Cordless Impact Wrench 1/2',     85,  225,  0.16),   # margin ~56%
        ('Portable Generator 3500W',      360,  820,  0.05),   # margin ~54%
        ('LED Work Light 60W',             33,   90,  0.18),   # margin ~55%
        ('Network Cable Tester',           48,  135,  0.07),   # margin ~66%
        ('Rotary Laser Level Kit',         90,  255,  0.14),   # margin ~59%
        ('Industrial Fan 24in',           128,  310,  0.28),   # margin ~42%
        ('Hydraulic Crimping Tool',        40,  112,  0.11),   # margin ~60%
        ('Power Inverter 2000W',          155,  390,  0.19),   # margin ~51%
        ('Digital Caliper 6in',            20,   58,  0.15),   # margin ~66%
        ('Voltage Regulator 1kVA',         72,  190,  0.12),   # margin ~57%
        ('Forklift Battery Charger',      385,  800,  0.44),   # margin ~14% below floor
        ('Air Impact Wrench 3/8in',        68,  185,  0.14),   # margin ~57%
        ('Dust Collector 1.5HP',          143,  348,  0.13),   # margin ~53%
        ('Soldering Station Digital',      46,  132,  0.06),   # margin ~63%
        ('Magnetic Drill Press 30mm',     248,  585,  0.35),   # margin ~35%
        ('Pneumatic Angle Grinder',        52,  150,  0.15),   # margin ~59%
        ('Infrared Thermometer -50-500C',  18,   52,  0.10),   # margin ~62%
        ('Industrial Humidifier 6gal',    172,  415,  0.14),   # margin ~51%
        ('Cable Management Tray 1U',       26,   76,  0.13),   # margin ~61%
        ('Electric Wire Rope Hoist',      310,  700,  0.08),   # margin ~52%
        ('Portable Welder MIG 220A',      175,  425,  0.18),   # margin ~51%
    ]

    normal_font = Font(name='Calibri', size=11)
    pct_fmt = '0%'
    currency_fmt = '$#,##0.00'

    for i, (product, cogs, list_price, disc) in enumerate(deals, 2):
        deal_id = f'DL-2025-{(i - 1):03d}'

        ws.cell(row=i, column=1, value=deal_id).font = normal_font
        ws.cell(row=i, column=2, value=product).font = normal_font

        cogs_cell = ws.cell(row=i, column=3, value=cogs)
        cogs_cell.font = normal_font
        cogs_cell.number_format = currency_fmt

        list_cell = ws.cell(row=i, column=4, value=list_price)
        list_cell.font = normal_font
        list_cell.number_format = currency_fmt

        disc_cell = ws.cell(row=i, column=5, value=disc)
        disc_cell.font = normal_font
        disc_cell.number_format = pct_fmt

        # Columns F-I intentionally empty (task asks to add these formulas)

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Rows: 41 (1 header + 40 data rows)')
    print(f'  Columns: A-I (F-I empty - no formulas, no conditional formatting, no comments)')


create_initial()
