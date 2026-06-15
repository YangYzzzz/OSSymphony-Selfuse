"""
Reward Script: Sales Pricing Margin Guard
Task ID: calc_sales_pricing_margin_guard_036
Domain: libreoffice_calc
Scoring:
  Component 1: Net Price formulas in F2:F41 (=D*(1-E))                 — 0.25 pts
  Component 2: Gross Profit (G=F-C) and Margin% (H=G/F) formulas      — 0.25 pts
  Component 3: Approval Status formulas in I2:I41 (nested IF)          — 0.20 pts
  Component 4: Conditional formatting on H2:H41 (red <0.10, orange <0.20) — 0.20 pts
  Component 5: At least one comment on critically low margin cell (H<0.10) — 0.10 pts
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_sales_pricing_margin_guard_036'


def normalize_formula(formula):
    """Normalize formula for comparison — remove spaces and uppercase."""
    if not formula:
        return ''
    return formula.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the sheet exists
    if 'DealPricer' not in wb.sheetnames:
        print("CRITICAL: Sheet 'DealPricer' not found in workbook")
        print(f"Sheets present: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['DealPricer']

    # Component 1: Net Price formulas in F2:F41 (=D*(1-E)) — 0.25 points
    # These should be =D2*(1-E2) patterns. Checks formula presence for all 40 rows.
    try:
        net_price_correct = 0
        net_price_total = 40
        for row in range(2, 42):  # rows 2-41
            cell_f = ws.cell(row=row, column=6)  # column F
            formula = cell_f.value
            if formula and isinstance(formula, str) and formula.startswith('='):
                # Check pattern: =Dx*(1-Ex) — normalize and check structure
                norm = normalize_formula(formula)
                # Expected: =D{row}*(1-E{row})
                expected_norm = normalize_formula(f'=D{row}*(1-E{row})')
                if norm == expected_norm:
                    net_price_correct += 1
                else:
                    # Accept equivalent patterns with different notation
                    # e.g. =D2*(1-E2) or =D2-D2*E2
                    if f'D{row}' in norm and f'E{row}' in norm and norm.startswith('='):
                        net_price_correct += 1

        ratio = net_price_correct / net_price_total
        if ratio >= 0.95:
            print(f"PASS: Component 1 — Net Price formulas: {net_price_correct}/{net_price_total} correct (0.25 pts)")
            total_score += 0.25
        elif ratio >= 0.5:
            partial = 0.25 * ratio
            print(f"PARTIAL: Component 1 — Net Price formulas: {net_price_correct}/{net_price_total} correct ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Net Price formulas: only {net_price_correct}/{net_price_total} correct (expected =D*(1-E) pattern)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Gross Profit (G=F-C) and Margin% (H=G/F) formulas — 0.25 points
    # G2:G41 should be =F-C pattern; H2:H41 should be =G/F pattern
    try:
        gp_correct = 0
        margin_correct = 0
        total_rows = 40
        for row in range(2, 42):
            # Check G (Gross Profit = Net Price - COGS)
            cell_g = ws.cell(row=row, column=7)  # column G
            formula_g = cell_g.value
            if formula_g and isinstance(formula_g, str) and formula_g.startswith('='):
                norm_g = normalize_formula(formula_g)
                expected_g = normalize_formula(f'=F{row}-C{row}')
                if norm_g == expected_g:
                    gp_correct += 1
                elif f'F{row}' in norm_g and f'C{row}' in norm_g:
                    gp_correct += 1

            # Check H (Margin % = Gross Profit / Net Price)
            cell_h = ws.cell(row=row, column=8)  # column H
            formula_h = cell_h.value
            if formula_h and isinstance(formula_h, str) and formula_h.startswith('='):
                norm_h = normalize_formula(formula_h)
                expected_h = normalize_formula(f'=G{row}/F{row}')
                if norm_h == expected_h:
                    margin_correct += 1
                elif f'G{row}' in norm_h and f'F{row}' in norm_h:
                    margin_correct += 1

        gp_ratio = gp_correct / total_rows
        margin_ratio = margin_correct / total_rows
        combined_ratio = (gp_ratio + margin_ratio) / 2.0

        if combined_ratio >= 0.95:
            print(f"PASS: Component 2 — GP formulas: {gp_correct}/{total_rows}, Margin% formulas: {margin_correct}/{total_rows} (0.25 pts)")
            total_score += 0.25
        elif combined_ratio >= 0.5:
            partial = 0.25 * combined_ratio
            print(f"PARTIAL: Component 2 — GP: {gp_correct}/{total_rows}, Margin%: {margin_correct}/{total_rows} ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — GP: {gp_correct}/{total_rows}, Margin%: {margin_correct}/{total_rows}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Approval Status formulas in I2:I41 (nested IF based on margin) — 0.20 points
    # =IF(H<0.10,"Manager Approval Required",IF(H<0.20,"Below Floor","Approved"))
    try:
        approval_correct = 0
        total_rows = 40
        for row in range(2, 42):
            cell_i = ws.cell(row=row, column=9)  # column I
            formula_i = cell_i.value
            if formula_i and isinstance(formula_i, str) and formula_i.startswith('='):
                norm_i = normalize_formula(formula_i)
                # Check key elements: IF, H{row}, 0.10, 0.20, Manager Approval Required, Below Floor, Approved
                has_if = 'IF(' in norm_i
                has_h_ref = f'H{row}' in norm_i
                has_0_10 = '0.10' in norm_i or '0.1' in norm_i
                has_0_20 = '0.20' in norm_i or '0.2' in norm_i
                has_mgr = 'MANAGERAPPROVALREQUIRED' in norm_i.replace(' ', '').replace('"', '')
                has_floor = 'BELOWFLOOR' in norm_i.replace(' ', '').replace('"', '')
                has_approved = 'APPROVED' in norm_i.replace(' ', '').replace('"', '')

                if has_if and has_h_ref and (has_0_10 or has_0_20) and has_mgr and has_floor and has_approved:
                    approval_correct += 1

        ratio = approval_correct / total_rows
        if ratio >= 0.95:
            print(f"PASS: Component 3 — Approval Status IF formulas: {approval_correct}/{total_rows} correct (0.20 pts)")
            total_score += 0.20
        elif ratio >= 0.5:
            partial = 0.20 * ratio
            print(f"PARTIAL: Component 3 — Approval Status IF formulas: {approval_correct}/{total_rows} ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Approval Status IF formulas: only {approval_correct}/{total_rows} correct")
            if approval_correct > 0:
                # Show first failing example
                for row in range(2, 42):
                    cell_i = ws.cell(row=row, column=9)
                    print(f"  Example I{row}: {repr(cell_i.value)}")
                    break
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Conditional formatting on H2:H41 — 0.20 points
    # Must have at least 2 rules: red fill for <0.10, orange fill for <0.20
    try:
        cf_rules = ws.conditional_formatting
        # Find rules that apply to H column range
        h_range_rules = []
        for cf_range in cf_rules:
            range_str = str(cf_range)
            if 'H' in range_str:
                for rule in cf_rules[cf_range]:
                    h_range_rules.append(rule)

        # Count distinct thresholds and fill colors found in H-range CF rules
        threshold_values = set()
        fill_color_set = set()

        for rule in h_range_rules:
            operator = getattr(rule, 'operator', None)
            formula = getattr(rule, 'formula', [])
            dxf = getattr(rule, 'dxf', None)

            if formula and dxf and hasattr(dxf, 'fill') and dxf.fill:
                # Collect threshold values
                for f in formula:
                    try:
                        val = float(str(f))
                        threshold_values.add(round(val, 2))
                    except (ValueError, TypeError):
                        pass
                # Collect fill colors
                try:
                    if dxf.fill.fgColor:
                        fill_color_set.add(dxf.fill.fgColor.rgb)
                except Exception:
                    pass

        # Need at least 2 distinct thresholds (0.10 and 0.20) and 2 distinct fill colors
        has_two_thresholds = len(threshold_values) >= 2
        has_two_colors = len(fill_color_set) >= 2
        # Also check that the thresholds include 0.1 and 0.2 specifically
        has_critical_threshold = any(abs(v - 0.10) < 0.01 for v in threshold_values)
        has_floor_threshold = any(abs(v - 0.20) < 0.01 for v in threshold_values)

        if has_two_thresholds and has_two_colors and has_critical_threshold and has_floor_threshold:
            print(f"PASS: Component 4 — Conditional formatting on H2:H41: 2 thresholds {sorted(threshold_values)}, 2 fill colors found (0.20 pts)")
            total_score += 0.20
        elif has_two_thresholds and has_two_colors:
            print(f"PARTIAL: Component 4 — CF has 2 thresholds and 2 colors but thresholds {sorted(threshold_values)} may not match 0.10/0.20 (0.10 pts)")
            total_score += 0.10
        elif len(h_range_rules) >= 1:
            print(f"PARTIAL: Component 4 — Some CF rules on H column but incomplete: thresholds={sorted(threshold_values)}, colors={fill_color_set} (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4 — No conditional formatting rules found on H column")
            print(f"  Total H-range rules found: {len(h_range_rules)}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: At least one comment on critically low margin cell (H value < 0.10) — 0.10 points
    # Comments should be on H column cells where margin is critically low
    try:
        h_comments = []
        for row in range(2, 42):
            cell_h = ws.cell(row=row, column=8)  # column H
            if cell_h.comment:
                comment_text = cell_h.comment.text if cell_h.comment.text else ''
                h_comments.append((f'H{row}', comment_text))

        if len(h_comments) >= 1:
            # Verify comment content references manager approval or warning
            valid_comments = []
            for coord, text in h_comments:
                text_upper = text.upper()
                if any(keyword in text_upper for keyword in ['MANAGER', 'APPROVAL', 'WARNING', 'CRITICAL', 'APPROVAL REQUIRED']):
                    valid_comments.append((coord, text))

            if len(valid_comments) >= 1:
                print(f"PASS: Component 5 — {len(valid_comments)} valid approval warning comment(s) found on H column cells: {[c[0] for c in valid_comments[:3]]} (0.10 pts)")
                total_score += 0.10
            elif len(h_comments) >= 1:
                # Comments exist but may not have specific keywords — give partial credit
                print(f"PARTIAL: Component 5 — {len(h_comments)} comment(s) on H column but may lack approval warning text (0.05 pts)")
                total_score += 0.05
        else:
            print(f"FAIL: Component 5 — No comments found on H column cells (expected at least 1 on margin < 0.10 cells)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
