"""
Initial Setup: Quarterly Report spreadsheet with quarter headers unmerged
Task ID: calc_cop_merge_005
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_cop_merge_005'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'QuarterlyReport'

    # --- Row 1: Quarter group headers (UNMERGED, per task description) ---
    ws['A1'] = 'Region'
    ws['B1'] = 'Q1'
    ws['C1'] = None  # empty
    ws['D1'] = None  # empty
    ws['E1'] = 'Q2'
    ws['F1'] = None  # empty
    ws['G1'] = None  # empty
    ws['H1'] = 'Q3'
    ws['I1'] = None  # empty
    ws['J1'] = None  # empty
    ws['K1'] = 'Q4'
    ws['L1'] = None  # empty
    ws['M1'] = None  # empty

    # Style row 1 headers (bold, light background)
    header_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    header_font = Font(bold=True, size=12)
    for col in range(1, 14):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Row 2: Month sub-headers ---
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    ws['A2'] = ''
    month_fill = PatternFill(start_color='FFE2EFDA', end_color='FFE2EFDA', fill_type='solid')
    month_font = Font(bold=True, size=11)
    for i, month in enumerate(months):
        cell = ws.cell(row=2, column=i + 2)  # B2:M2
        cell.value = month
        cell.font = month_font
        cell.fill = month_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Data rows: Realistic regional sales data ---
    regions = [
        'North America',
        'Europe',
        'Asia Pacific',
        'Latin America',
        'Middle East',
        'Africa',
        'South Asia',
        'Oceania',
        'Central Asia',
        'Caribbean',
        'Nordic',
        'Benelux',
    ]

    # Monthly revenue data per region (realistic business numbers, in thousands)
    data = [
        [3420, 3180, 3670, 3540, 3820, 3990, 4120, 4250, 3980, 4300, 4560, 4780],
        [2850, 2760, 3010, 2940, 3120, 3280, 3350, 3410, 3190, 3450, 3620, 3840],
        [1980, 1870, 2140, 2080, 2310, 2450, 2560, 2640, 2490, 2710, 2890, 3050],
        [870,  820,  910,  880,  950,  1020, 1080, 1110, 1040, 1130, 1200, 1280],
        [640,  600,  680,  660,  710,  760,  800,  830,  780,  850,  900,  960],
        [430,  410,  460,  440,  480,  510,  540,  560,  520,  570,  600,  640],
        [1240, 1180, 1310, 1270, 1380, 1460, 1530, 1580, 1490, 1620, 1720, 1830],
        [560,  530,  590,  570,  620,  660,  690,  710,  680,  730,  780,  830],
        [320,  300,  340,  330,  360,  380,  400,  410,  390,  420,  450,  480],
        [280,  260,  300,  290,  310,  330,  350,  360,  340,  370,  390,  420],
        [920,  880,  970,  940,  1020, 1080, 1130, 1170, 1100, 1200, 1270, 1350],
        [750,  710,  790,  760,  830,  880,  920,  950,  900,  970,  1030, 1100],
    ]

    data_font = Font(size=11)
    for r_idx, (region, row_data) in enumerate(zip(regions, data), start=3):
        ws.cell(row=r_idx, column=1, value=region).font = data_font
        ws.cell(row=r_idx, column=1).alignment = Alignment(horizontal='left', vertical='center')
        for c_idx, val in enumerate(row_data, start=2):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = data_font
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.number_format = '#,##0'

    # --- Column widths ---
    ws.column_dimensions['A'].width = 18
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
        ws.column_dimensions[col_letter].width = 9

    # --- Row heights ---
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet: QuarterlyReport')
    print(f'Rows: {ws.max_row}, Columns: {ws.max_column}')
    print(f'No merged cells in initial file')


create_initial()
