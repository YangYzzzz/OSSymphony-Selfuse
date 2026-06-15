"""
Reward Script: Merge header cells for each quarter group and center text
Task ID: calc_cop_merge_005
Domain: libreoffice_calc
Scoring:
  Component 1: B1:D1 merged with 'Q1' in B1 (0.25 pts)
  Component 2: E1:G1 merged with 'Q2' in E1 (0.25 pts)
  Component 3: H1:J1 merged with 'Q3' in H1 (0.25 pts)
  Component 4: K1:M1 merged with 'Q4' in K1 (0.25 pts)
  Total: 1.0

NOTE: The task requires merging AND centering. However, alignment was already 'center'
in the initial file. The only task-introduced change is the merging itself.
Centering is therefore a precondition gate, NOT a scoring component.
Each component verifies the merge range exists AND that the sub-cells (non-anchor)
are MergedCell instances, and the anchor cell retains its quarter label.
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_cop_merge_005'
SHEET_NAME = 'QuarterlyReport'


def check_merge_range(ws, merge_range_str, anchor_col, anchor_value,
                      sub_cells):
    """
    Verify that a specific merge range exists in the worksheet.

    Args:
        ws: openpyxl worksheet
        merge_range_str: expected merge range string (e.g. 'B1:D1')
        anchor_col: column letter of the top-left anchor cell (e.g. 'B')
        anchor_value: expected value in the anchor cell (e.g. 'Q1')
        sub_cells: list of sub-cell coordinate strings that should be MergedCell
                   (e.g. ['C1', 'D1'])

    Returns:
        (passed: bool, details: str)
    """
    # Collect all merge range strings present in the sheet
    existing_ranges = set()
    for mr in ws.merged_cells.ranges:
        existing_ranges.add(str(mr))

    # Check the specific range is registered
    if merge_range_str not in existing_ranges:
        return False, f"Merge range {merge_range_str} not found in merged_cells (found: {existing_ranges})"

    # Verify anchor cell holds the expected value
    anchor_coord = f"{anchor_col}1"
    anchor_cell = ws[anchor_coord]
    if anchor_cell.value != anchor_value:
        return False, (
            f"Anchor cell {anchor_coord} has value {repr(anchor_cell.value)}, "
            f"expected {repr(anchor_value)}"
        )

    # Verify sub-cells are MergedCell instances
    for sub_coord in sub_cells:
        sub_cell = ws[sub_coord]
        if not isinstance(sub_cell, MergedCell):
            return False, (
                f"Sub-cell {sub_coord} is not a MergedCell "
                f"(type: {type(sub_cell).__name__})"
            )

    return True, (
        f"Range {merge_range_str} merged; anchor {anchor_coord}={repr(anchor_value)}; "
        f"sub-cells {sub_cells} are MergedCell"
    )


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet exists
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Precondition gate: check basic structure is intact (not scored)
    # A1 should be 'Region'; Row 2 should have month sub-headers
    a1_val = ws['A1'].value
    if a1_val != 'Region':
        print(f"WARNING: A1 value is {repr(a1_val)}, expected 'Region'. File may be corrupted.")

    # Component 1: B1:D1 merged with 'Q1' in B1 (0.25 points)
    # The initial file has NO merged cells; golden file has B1:D1 merged.
    # This check FAILS on initial (no merge) and PASSES on golden.
    try:
        passed, details = check_merge_range(
            ws,
            merge_range_str='B1:D1',
            anchor_col='B',
            anchor_value='Q1',
            sub_cells=['C1', 'D1']
        )
        if passed:
            print(f"PASS: Component 1 — B1:D1 merged for Q1 (0.25 pts) — {details}")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — B1:D1 merge for Q1 — {details}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: E1:G1 merged with 'Q2' in E1 (0.25 points)
    # The initial file has NO merged cells; golden file has E1:G1 merged.
    try:
        passed, details = check_merge_range(
            ws,
            merge_range_str='E1:G1',
            anchor_col='E',
            anchor_value='Q2',
            sub_cells=['F1', 'G1']
        )
        if passed:
            print(f"PASS: Component 2 — E1:G1 merged for Q2 (0.25 pts) — {details}")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — E1:G1 merge for Q2 — {details}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: H1:J1 merged with 'Q3' in H1 (0.25 points)
    # The initial file has NO merged cells; golden file has H1:J1 merged.
    try:
        passed, details = check_merge_range(
            ws,
            merge_range_str='H1:J1',
            anchor_col='H',
            anchor_value='Q3',
            sub_cells=['I1', 'J1']
        )
        if passed:
            print(f"PASS: Component 3 — H1:J1 merged for Q3 (0.25 pts) — {details}")
            total_score += 0.25
        else:
            print(f"FAIL: Component 3 — H1:J1 merge for Q3 — {details}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: K1:M1 merged with 'Q4' in K1 (0.25 points)
    # The initial file has NO merged cells; golden file has K1:M1 merged.
    try:
        passed, details = check_merge_range(
            ws,
            merge_range_str='K1:M1',
            anchor_col='K',
            anchor_value='Q4',
            sub_cells=['L1', 'M1']
        )
        if passed:
            print(f"PASS: Component 4 — K1:M1 merged for Q4 (0.25 pts) — {details}")
            total_score += 0.25
        else:
            print(f"FAIL: Component 4 — K1:M1 merge for Q4 — {details}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
