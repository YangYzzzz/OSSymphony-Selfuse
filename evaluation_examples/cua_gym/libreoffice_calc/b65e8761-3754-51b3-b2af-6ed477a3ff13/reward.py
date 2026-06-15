"""
Reward Script: Employee Skills Assessment Matrix
Task ID: calc_grs_078
Domain: libreoffice_calc
Scoring:
  Component 1: Category average columns added to Skills Inventory (0.20)
  Component 2: Color scale conditional formatting on skill cells (0.15)
  Component 3: Gap column added to Skills Gap Analysis (0.15)
  Component 4: Conditional formatting on Gap column (red for gaps) (0.10)
  Component 5: Radar Charts sheet with 8 radar charts (0.25)
  Component 6: Team Heatmap sheet exists with data and fill colors (0.15)
"""

import os
import openpyxl
from openpyxl.chart.radar_chart import RadarChart

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_078'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    sheet_names = wb.sheetnames

    # ----------------------------------------------------------------
    # Component 1: Category average columns in Skills Inventory (0.20)
    # Golden adds 5 avg columns (V-Z): Avg: Technical, Software, Communication, Leadership, Domain
    # These contain AVERAGE formulas. Initial has only 21 cols (A-U).
    # ----------------------------------------------------------------
    try:
        if 'Skills Inventory' not in sheet_names:
            print("FAIL: Component 1 — 'Skills Inventory' sheet missing")
        else:
            ws = wb['Skills Inventory']
            # Check that there are columns beyond U (col 21) with average headers
            avg_col_count = 0
            expected_avg_keywords = ['technical', 'software', 'communication', 'leadership', 'domain']
            found_keywords = []
            for c in range(22, ws.max_column + 1):
                header = ws.cell(1, c).value
                if header and 'avg' in str(header).lower():
                    avg_col_count += 1
                    for kw in expected_avg_keywords:
                        if kw in str(header).lower():
                            found_keywords.append(kw)

            # Also check that at least one cell in row 3 of these avg cols has a formula
            has_formula = False
            for c in range(22, ws.max_column + 1):
                val = ws.cell(3, c).value
                if val and isinstance(val, str) and '=AVERAGE' in val.upper():
                    has_formula = True
                    break

            if avg_col_count >= 5 and has_formula:
                print(f"PASS: Component 1 — {avg_col_count} average columns found with AVERAGE formulas (0.20 pts)")
                total_score += 0.20
            elif avg_col_count >= 3 and has_formula:
                print(f"PARTIAL: Component 1 — {avg_col_count}/5 average columns found (0.10 pts)")
                total_score += 0.10
            elif avg_col_count >= 1:
                print(f"PARTIAL: Component 1 — {avg_col_count}/5 average columns found (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 1 — No average columns found beyond col 21 (max_col={ws.max_column})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ----------------------------------------------------------------
    # Component 2: Color scale conditional formatting on skill cells (0.15)
    # Golden has colorScale CF on B3:U10 in Skills Inventory.
    # Initial has NO CF at all.
    # ----------------------------------------------------------------
    try:
        if 'Skills Inventory' not in sheet_names:
            print("FAIL: Component 2 — 'Skills Inventory' sheet missing")
        else:
            ws = wb['Skills Inventory']
            color_scale_found = False
            for cf in ws.conditional_formatting:
                for rule in cf.rules:
                    if rule.type == 'colorScale':
                        color_scale_found = True
                        break
                if color_scale_found:
                    break

            if color_scale_found:
                print(f"PASS: Component 2 — Color scale conditional formatting found on Skills Inventory (0.15 pts)")
                total_score += 0.15
            else:
                print("FAIL: Component 2 — No color scale CF found on Skills Inventory")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ----------------------------------------------------------------
    # Component 3: Gap column in Skills Gap Analysis (0.15)
    # Golden has col F = "Gap" with computed gap values. Initial has only 5 cols (A-E).
    # ----------------------------------------------------------------
    try:
        if 'Skills Gap Analysis' not in sheet_names:
            print("FAIL: Component 3 — 'Skills Gap Analysis' sheet missing")
        else:
            ws2 = wb['Skills Gap Analysis']
            # Check if there's a 6th column with gap-related header
            gap_col = None
            for c in range(1, ws2.max_column + 1):
                h = ws2.cell(1, c).value
                if h and 'gap' in str(h).lower():
                    gap_col = c
                    break

            if gap_col is not None:
                # Verify at least some gap values exist (non-None in data rows)
                gap_values_count = 0
                for r in range(2, min(20, ws2.max_row + 1)):
                    val = ws2.cell(r, gap_col).value
                    if val is not None:
                        gap_values_count += 1

                if gap_values_count >= 5:
                    print(f"PASS: Component 3 — Gap column found at col {gap_col} with {gap_values_count} values (0.15 pts)")
                    total_score += 0.15
                else:
                    print(f"PARTIAL: Component 3 — Gap column header found but only {gap_values_count} values (0.05 pts)")
                    total_score += 0.05
            else:
                print(f"FAIL: Component 3 — No 'Gap' column found (max_col={ws2.max_column})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ----------------------------------------------------------------
    # Component 4: Conditional formatting on Gap column (red for gaps) (0.10)
    # Golden has cellIs CF on F2:F161. Initial has NO CF.
    # ----------------------------------------------------------------
    try:
        if 'Skills Gap Analysis' not in sheet_names:
            print("FAIL: Component 4 — 'Skills Gap Analysis' sheet missing")
        else:
            ws2 = wb['Skills Gap Analysis']
            gap_cf_found = False
            for cf in ws2.conditional_formatting:
                for rule in cf.rules:
                    if rule.type == 'cellIs':
                        gap_cf_found = True
                        break
                if gap_cf_found:
                    break

            if gap_cf_found:
                print(f"PASS: Component 4 — Conditional formatting on Gap column found (0.10 pts)")
                total_score += 0.10
            else:
                print("FAIL: Component 4 — No cellIs CF found on Skills Gap Analysis")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ----------------------------------------------------------------
    # Component 5: Radar Charts sheet with radar charts (0.25)
    # Golden has "Radar Charts" sheet with 8 RadarChart objects (one per employee).
    # Initial has NO such sheet.
    # ----------------------------------------------------------------
    try:
        radar_sheet_exists = False
        radar_sheet_name = None
        for sn in sheet_names:
            if 'radar' in sn.lower():
                radar_sheet_exists = True
                radar_sheet_name = sn
                break

        if not radar_sheet_exists:
            print("FAIL: Component 5 — No radar chart sheet found")
        else:
            ws3 = wb[radar_sheet_name]
            num_charts = len(ws3._charts)
            # Count how many are radar charts specifically
            radar_count = 0
            for ch in ws3._charts:
                if isinstance(ch, RadarChart):
                    radar_count += 1

            if radar_count >= 8:
                print(f"PASS: Component 5 — {radar_count} radar charts found on '{radar_sheet_name}' (0.25 pts)")
                total_score += 0.25
            elif radar_count >= 4:
                print(f"PARTIAL: Component 5 — {radar_count}/8 radar charts found (0.15 pts)")
                total_score += 0.15
            elif radar_count >= 1:
                print(f"PARTIAL: Component 5 — {radar_count}/8 radar charts found (0.08 pts)")
                total_score += 0.08
            elif num_charts >= 1:
                # Charts exist but not radar type — partial credit
                print(f"PARTIAL: Component 5 — {num_charts} charts found but {radar_count} are radar type (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 5 — No charts on '{radar_sheet_name}' sheet")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # ----------------------------------------------------------------
    # Component 6: Team Heatmap sheet with data and fill colors (0.15)
    # Golden has "Team Heatmap" sheet with employee skill data and colored cells.
    # Initial has NO such sheet.
    # ----------------------------------------------------------------
    try:
        heatmap_sheet_exists = False
        heatmap_sheet_name = None
        for sn in sheet_names:
            if 'heatmap' in sn.lower():
                heatmap_sheet_exists = True
                heatmap_sheet_name = sn
                break

        if not heatmap_sheet_exists:
            print("FAIL: Component 6 — No heatmap sheet found")
        else:
            ws4 = wb[heatmap_sheet_name]

            # Check that the sheet has data (at least a few rows of employee skills)
            has_data = ws4.max_row >= 5 and ws4.max_column >= 5

            # Check for fill colors on data cells (heatmap coloring)
            # In golden, cells have fills like FF63BE7B
            colored_cells = 0
            for r in range(2, min(ws4.max_row + 1, 15)):
                for c in range(2, min(ws4.max_column + 1, 10)):
                    cell = ws4.cell(r, c)
                    try:
                        fill_rgb = cell.fill.fgColor.rgb
                        if fill_rgb and fill_rgb != '00000000' and cell.fill.patternType == 'solid':
                            colored_cells += 1
                    except Exception:
                        pass

            # Also check for CF-based heatmap (alternative implementation)
            has_cf = False
            for cf in ws4.conditional_formatting:
                for rule in cf.rules:
                    if rule.type in ('colorScale', 'cellIs'):
                        has_cf = True
                        break

            if has_data and (colored_cells >= 10 or has_cf):
                print(f"PASS: Component 6 — Heatmap sheet '{heatmap_sheet_name}' has data and coloring "
                      f"(colored_cells={colored_cells}, has_cf={has_cf}) (0.15 pts)")
                total_score += 0.15
            elif has_data:
                print(f"PARTIAL: Component 6 — Heatmap sheet exists with data but limited coloring "
                      f"(colored={colored_cells}, cf={has_cf}) (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 6 — Heatmap sheet exists but insufficient data "
                      f"(rows={ws4.max_row}, cols={ws4.max_column})")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
