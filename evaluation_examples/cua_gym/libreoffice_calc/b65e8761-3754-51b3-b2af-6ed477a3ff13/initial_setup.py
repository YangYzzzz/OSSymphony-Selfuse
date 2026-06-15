"""
Initial Setup: Employee Skills Assessment Matrix
Task ID: calc_grs_078
Domain: libreoffice_calc

Creates an initial spreadsheet with:
- Skills Inventory sheet: 8 employees x 20 skills with ratings 1-4
- Skills Gap Analysis sheet: basic structure with headers only (no comparisons yet)
No conditional formatting, no category averages, no charts, no gap highlighting.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_078'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


# ---------- Skill definitions by category ----------
TECHNICAL_SKILLS = [
    "Python Programming", "Cloud Architecture", "Data Engineering",
    "API Design", "DevOps/CI-CD"
]
SOFTWARE_TOOLS = [
    "AWS Console", "Docker/Kubernetes", "Git/GitHub",
    "Jira/Confluence", "SQL Databases"
]
COMMUNICATION = [
    "Technical Writing", "Stakeholder Presentations", "Cross-team Collaboration"
]
LEADERSHIP = [
    "Mentoring", "Project Management", "Decision Making"
]
DOMAIN_KNOWLEDGE = [
    "Financial Systems", "Healthcare Compliance", "E-commerce Platforms", "Data Privacy/GDPR"
]

ALL_SKILLS = TECHNICAL_SKILLS + SOFTWARE_TOOLS + COMMUNICATION + LEADERSHIP + DOMAIN_KNOWLEDGE

EMPLOYEES = [
    "Sarah Chen",
    "Marcus Johnson",
    "Priya Patel",
    "David Kim",
    "Elena Rodriguez",
    "James Okafor",
    "Aisha Nakamura",
    "Thomas Bergmann"
]

ROLES = [
    "Senior Backend Engineer",
    "DevOps Lead",
    "Data Engineer",
    "Full Stack Developer",
    "Cloud Architect",
    "QA Engineer",
    "ML Engineer",
    "Frontend Developer"
]

# Skill ratings (1-4) for each employee across 20 skills
# Designed to be realistic with strengths/weaknesses per role
RATINGS = [
    # Sarah Chen - Senior Backend Engineer
    [4, 3, 3, 4, 3,  3, 3, 4, 3, 4,  3, 3, 3,  2, 3, 3,  2, 1, 3, 2],
    # Marcus Johnson - DevOps Lead
    [3, 4, 2, 3, 4,  4, 4, 4, 3, 3,  2, 3, 3,  3, 4, 3,  1, 1, 2, 2],
    # Priya Patel - Data Engineer
    [4, 2, 4, 3, 3,  3, 3, 3, 2, 4,  3, 2, 2,  2, 2, 2,  3, 2, 2, 3],
    # David Kim - Full Stack Developer
    [3, 2, 2, 3, 2,  2, 2, 4, 3, 3,  2, 3, 3,  1, 2, 2,  1, 1, 4, 1],
    # Elena Rodriguez - Cloud Architect
    [3, 4, 3, 4, 3,  4, 4, 3, 3, 3,  3, 4, 3,  3, 3, 4,  2, 2, 3, 3],
    # James Okafor - QA Engineer
    [3, 2, 2, 2, 3,  2, 2, 3, 4, 2,  3, 2, 3,  2, 2, 2,  1, 3, 2, 2],
    # Aisha Nakamura - ML Engineer
    [4, 3, 4, 3, 2,  3, 3, 3, 2, 4,  2, 2, 2,  2, 2, 3,  2, 2, 1, 4],
    # Thomas Bergmann - Frontend Developer
    [2, 2, 1, 2, 2,  1, 2, 4, 3, 1,  3, 3, 4,  1, 2, 2,  1, 1, 3, 1],
]

# Required levels per role for gap analysis (same 20 skills)
REQUIRED_LEVELS = [
    # Senior Backend Engineer
    [4, 3, 3, 4, 3,  3, 3, 4, 3, 4,  3, 3, 3,  3, 3, 3,  2, 2, 3, 3],
    # DevOps Lead
    [3, 4, 3, 3, 4,  4, 4, 4, 3, 3,  3, 3, 3,  3, 4, 4,  2, 2, 2, 3],
    # Data Engineer
    [4, 3, 4, 3, 3,  4, 3, 3, 3, 4,  3, 3, 3,  2, 3, 3,  3, 2, 2, 4],
    # Full Stack Developer
    [3, 3, 2, 4, 3,  3, 3, 4, 3, 3,  3, 3, 3,  2, 3, 3,  2, 2, 4, 2],
    # Cloud Architect
    [3, 4, 3, 4, 4,  4, 4, 3, 3, 3,  3, 4, 4,  3, 4, 4,  2, 2, 3, 3],
    # QA Engineer
    [3, 2, 2, 3, 3,  2, 3, 3, 4, 3,  3, 3, 3,  2, 3, 3,  2, 3, 2, 3],
    # ML Engineer
    [4, 3, 4, 3, 3,  3, 3, 3, 3, 4,  3, 3, 3,  2, 3, 3,  3, 2, 2, 4],
    # Frontend Developer
    [2, 2, 1, 3, 2,  2, 2, 4, 3, 2,  3, 4, 4,  2, 3, 3,  1, 1, 4, 2],
]


def create_initial():
    wb = openpyxl.Workbook()

    # ===== Sheet 1: Skills Inventory =====
    ws1 = wb.active
    ws1.title = "Skills Inventory"

    # Header row
    headers = ["Employee"] + ALL_SKILLS
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Category sub-header row (row 2)
    cat_row_font = Font(name="Calibri", size=9, italic=True, color="333333")
    cat_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
    ws1.cell(row=2, column=1, value="Role").font = cat_row_font
    ws1.cell(row=2, column=1).fill = cat_fill
    ws1.cell(row=2, column=1).alignment = Alignment(horizontal="center")
    ws1.cell(row=2, column=1).border = thin_border

    # Fill category labels across skill columns
    cat_mapping = (
        [("Technical Skills", len(TECHNICAL_SKILLS))] +
        [("Software Tools", len(SOFTWARE_TOOLS))] +
        [("Communication", len(COMMUNICATION))] +
        [("Leadership", len(LEADERSHIP))] +
        [("Domain Knowledge", len(DOMAIN_KNOWLEDGE))]
    )
    col_offset = 2
    for cat_name, cat_count in cat_mapping:
        for i in range(cat_count):
            cell = ws1.cell(row=2, column=col_offset + i, value=cat_name)
            cell.font = cat_row_font
            cell.fill = cat_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        col_offset += cat_count

    # Employee data rows (rows 3-10)
    data_align = Alignment(horizontal="center", vertical="center")
    name_align = Alignment(horizontal="left", vertical="center")

    for emp_idx, emp_name in enumerate(EMPLOYEES):
        row = emp_idx + 3
        # Employee name
        cell = ws1.cell(row=row, column=1, value=emp_name)
        cell.font = Font(name="Calibri", size=11)
        cell.alignment = name_align
        cell.border = thin_border

        # Skill ratings
        for skill_idx, rating in enumerate(RATINGS[emp_idx]):
            cell = ws1.cell(row=row, column=skill_idx + 2, value=rating)
            cell.alignment = data_align
            cell.border = thin_border
            cell.number_format = '0'

    # Column widths
    ws1.column_dimensions["A"].width = 22
    for col_idx in range(2, len(ALL_SKILLS) + 2):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws1.column_dimensions[col_letter].width = 14

    # Row heights
    ws1.row_dimensions[1].height = 45
    ws1.row_dimensions[2].height = 20

    # Freeze panes
    ws1.freeze_panes = "B3"

    # Rating scale legend (below data)
    legend_row = 12
    ws1.cell(row=legend_row, column=1, value="Rating Scale:").font = Font(bold=True, size=10)
    ws1.cell(row=legend_row + 1, column=1, value="1 = No Knowledge")
    ws1.cell(row=legend_row + 2, column=1, value="2 = Basic")
    ws1.cell(row=legend_row + 3, column=1, value="3 = Proficient")
    ws1.cell(row=legend_row + 4, column=1, value="4 = Expert")

    # ===== Sheet 2: Skills Gap Analysis =====
    ws2 = wb.create_sheet("Skills Gap Analysis")

    # Basic structure - headers only, no gap calculations yet
    gap_headers = ["Employee", "Role", "Skill", "Current Level", "Required Level"]
    for col_idx, header in enumerate(gap_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Populate raw data rows (employee + role + skill + current level + required level)
    # No gap column, no red highlighting
    data_row = 2
    for emp_idx, emp_name in enumerate(EMPLOYEES):
        for skill_idx, skill_name in enumerate(ALL_SKILLS):
            ws2.cell(row=data_row, column=1, value=emp_name).border = thin_border
            ws2.cell(row=data_row, column=2, value=ROLES[emp_idx]).border = thin_border
            ws2.cell(row=data_row, column=3, value=skill_name).border = thin_border
            ws2.cell(row=data_row, column=4, value=RATINGS[emp_idx][skill_idx]).border = thin_border
            ws2.cell(row=data_row, column=4).alignment = data_align
            ws2.cell(row=data_row, column=5, value=REQUIRED_LEVELS[emp_idx][skill_idx]).border = thin_border
            ws2.cell(row=data_row, column=5).alignment = data_align
            data_row += 1

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 26
    ws2.column_dimensions["C"].width = 28
    ws2.column_dimensions["D"].width = 16
    ws2.column_dimensions["E"].width = 16
    ws2.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
