"""
Reward Script: Add Total Row with SUM Formulas and Multi-Series Line Chart
Task ID: osworld_calc_total_row_line_chart_007
Domain: libreoffice_calc
Scoring:
  Component 1: Total row label "Total" in column A (0.2 pts)
  Component 2: SUM formulas in Total row for all 11 month columns (0.4 pts)
  Component 3: Line chart exists in the sheet (0.2 pts)
  Component 4: Chart has Total series referencing row 6 (0.1 pts)
  Component 5: Chart has all 4 individual product series (0.1 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_total_row_line_chart_007'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Sales' sheet must exist
    if 'Sales' not in wb.sheetnames:
        print("CRITICAL: 'Sales' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Sales']

    # Determine the last data row (Total row must be after the 4 product rows)
    # Initial file has rows 1-5 (header + 4 products). Total row is expected at row 6.
    # We'll find the Total row by scanning up to row 10.
    total_row_idx = None
    for r in range(2, 11):
        cell_val = ws.cell(row=r, column=1).value
        if cell_val is not None and str(cell_val).strip().lower() == 'total':
            total_row_idx = r
            break

    # Component 1: Total row with label "Total" in column A (0.2 pts)
    try:
        if total_row_idx is not None:
            label_cell = ws.cell(row=total_row_idx, column=1).value
            print(f"PASS: Component 1 — 'Total' label found in A{total_row_idx} (value: '{label_cell}') (0.2 pts)")
            total_score += 0.2
        else:
            print("FAIL: Component 1 — No 'Total' label found in column A (rows 2-10)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: SUM formulas in Total row for all 11 month columns B-L (0.4 pts)
    # Each correct SUM formula earns 0.4/11 ~ 0.036 pts, total capped at 0.4
    try:
        if total_row_idx is not None:
            # Expected column letters for 11 months: B(2) to L(12)
            # Each formula should be =SUM(Xn:Xm) where n=2 and m = total_row_idx - 1
            month_cols = range(2, 13)  # columns 2-12 (B to L)
            sum_formula_count = 0
            for col_idx in month_cols:
                cell = ws.cell(row=total_row_idx, column=col_idx)
                val = cell.value
                if val is not None and isinstance(val, str):
                    # Normalize: remove spaces, uppercase
                    normalized = val.upper().replace(' ', '')
                    # Accept any SUM formula that references the column and covers rows 2-5 area
                    if normalized.startswith('=SUM(') and normalized.endswith(')'):
                        sum_formula_count += 1
                    else:
                        print(f"  WARN: Column {col_idx} has value '{val}' (not a SUM formula)")
                else:
                    print(f"  WARN: Column {col_idx} has value '{val}' (expected SUM formula string)")

            if sum_formula_count == 11:
                print(f"PASS: Component 2 — All 11 SUM formulas found in Total row (0.4 pts)")
                total_score += 0.4
            elif sum_formula_count >= 6:
                partial = round(0.4 * sum_formula_count / 11, 4)
                print(f"PARTIAL: Component 2 — {sum_formula_count}/11 SUM formulas found ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 2 — Only {sum_formula_count}/11 SUM formulas found in Total row")
        else:
            print("FAIL: Component 2 — Cannot check SUM formulas because Total row not found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: A line chart exists in the sheet (0.2 pts)
    try:
        charts = ws._charts
        line_charts = [c for c in charts if type(c).__name__ == 'LineChart']
        if len(line_charts) >= 1:
            print(f"PASS: Component 3 — Line chart found (total charts: {len(charts)}, line charts: {len(line_charts)}) (0.2 pts)")
            total_score += 0.2
        elif len(charts) >= 1:
            # Some chart exists but is not a line chart
            print(f"FAIL: Component 3 — Chart exists but is not a LineChart (types: {[type(c).__name__ for c in charts]})")
        else:
            print("FAIL: Component 3 — No chart found in the sheet")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Chart has a series referencing the Total row (row 6 / total_row_idx) (0.1 pts)
    try:
        charts = ws._charts
        line_charts = [c for c in charts if type(c).__name__ == 'LineChart']
        if len(line_charts) >= 1:
            chart = line_charts[0]

            def _series_refs_total(ser, total_row):
                """Return True if this series references the Total row."""
                if ser.val and ser.val.numRef:
                    ref = ser.val.numRef.f
                    row_marker = f'$B${total_row}:$L${total_row}'
                    if row_marker in ref:
                        return True
                # Check series title
                try:
                    if ser.title and hasattr(ser.title, 'v') and ser.title.v == 'Total':
                        return True
                except Exception:
                    pass
                return False

            total_series_count = sum(
                1 for ser in chart.series
                if total_row_idx is not None and _series_refs_total(ser, total_row_idx)
            )
            if total_series_count >= 1:
                print(f"PASS: Component 4 — Chart contains Total series referencing row {total_row_idx} (0.1 pts)")
                total_score += 0.1
            else:
                # Provide debug info
                series_refs = [
                    ser.val.numRef.f if (ser.val and ser.val.numRef) else None
                    for ser in chart.series
                ]
                print(f"FAIL: Component 4 — No series found referencing Total row. Series refs: {series_refs}")
        else:
            print("FAIL: Component 4 — No line chart to check for Total series")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Chart has all 4 individual product series (0.1 pts)
    # Each product row is rows 2-5 in the initial data
    try:
        charts = ws._charts
        line_charts = [c for c in charts if type(c).__name__ == 'LineChart']
        if len(line_charts) >= 1:
            chart = line_charts[0]
            # Count series referencing individual product rows (rows 2-5)
            product_rows_found = set()
            for ser in chart.series:
                if ser.val and ser.val.numRef:
                    ref = ser.val.numRef.f
                    for prod_row in range(2, 6):  # rows 2, 3, 4, 5
                        row_marker = f'$B${prod_row}:$L${prod_row}'
                        row_marker_alt = f'B{prod_row}:L{prod_row}'
                        if row_marker in ref or row_marker_alt in ref.upper():
                            product_rows_found.add(prod_row)
            if len(product_rows_found) >= 4:
                print(f"PASS: Component 5 — All 4 product series found in chart (rows: {sorted(product_rows_found)}) (0.1 pts)")
                total_score += 0.1
            elif len(product_rows_found) >= 1:
                partial = round(0.1 * len(product_rows_found) / 4, 4)
                print(f"PARTIAL: Component 5 — {len(product_rows_found)}/4 product series found ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 5 — No individual product series found in chart (total series: {len(chart.series)})")
        else:
            print("FAIL: Component 5 — No line chart to check for product series")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {round(final_score, 4)}")
    return final_score


# Default: test against canonical artifact path in the VM environment
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
