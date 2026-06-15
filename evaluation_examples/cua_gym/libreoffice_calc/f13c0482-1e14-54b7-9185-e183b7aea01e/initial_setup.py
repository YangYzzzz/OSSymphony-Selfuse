"""
Initial Setup: Product sales spreadsheet with monthly data for 4 products.
Task ID: osworld_calc_total_row_line_chart_007
Domain: libreoffice_calc

Creates a spreadsheet with:
- Column A: Product names (4 products)
- Columns B-L: Monthly sales data (Jan-Nov, 11 months)
- Row 1: Headers
- Rows 2-5: 4 product data rows
- NO Total row (task requires adding it)
- NO chart (task requires creating it)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_total_row_line_chart_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Sales ---
    ws = wb.active
    ws.title = 'Sales'

    # Month names for 11 months (Jan - Nov)
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov']

    # Headers: Row 1
    ws.cell(row=1, column=1, value='Product')
    for col_idx, month in enumerate(months, 2):
        ws.cell(row=1, column=col_idx, value=month)

    # Style headers: bold font
    header_font = Font(bold=True, name='Calibri', size=11)
    header_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    for col_idx in range(1, 13):  # A to L
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Product sales data - realistic business numbers
    # 4 products with 11 months of data
    products_data = [
        # Product name, Jan-Nov monthly sales
        ('UltraSlim Laptop Pro', [42350, 38920, 51200, 47800, 55600, 62100,
                                   58750, 63400, 71200, 68900, 76500]),
        ('ProSound Wireless Earbuds', [18900, 22450, 19800, 24300, 27600, 31200,
                                        29800, 33500, 36100, 34200, 38900]),
        ('SmartHome Hub 3.0', [9800, 11200, 10500, 13400, 15800, 17200,
                                16400, 19500, 21300, 20100, 23500]),
        ('ErgoDesk Standing Converter', [6200, 7100, 8400, 9200, 10800, 12400,
                                          11500, 13800, 15200, 14600, 16800]),
    ]

    for row_offset, (product_name, monthly_sales) in enumerate(products_data, 2):
        ws.cell(row=row_offset, column=1, value=product_name)
        for col_idx, sales_val in enumerate(monthly_sales, 2):
            ws.cell(row=row_offset, column=col_idx, value=sales_val)

    # Style product name column: slightly wider
    ws.column_dimensions['A'].width = 32
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        ws.column_dimensions[col_letter].width = 10

    # Number format for sales columns
    for row_idx in range(2, 6):
        for col_idx in range(2, 13):
            ws.cell(row=row_idx, column=col_idx).number_format = '#,##0'

    # Row height for header
    ws.row_dimensions[1].height = 20

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: Open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
