"""
Initial Setup: Pivot table with empty cells (no zeros) for calc_pivot_086
Task ID: calc_pivot_086
Domain: libreoffice_calc

Creates a workbook with:
- "Data" sheet: raw sales data by Product, Region
- "PivotSheet": a pivot-table-style layout with Product as rows, Region as columns,
  SUM of Sales as values. Some Product/Region combos have no data -> blank cells.
  Grand total across all non-empty cells = 165000.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_086'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ========== Pivot table values ==========
    # Products x Regions, some combos blank, grand total = 165000
    products = ['Widget', 'Gadget', 'Gizmo', 'Doohickey', 'Thingamajig']
    regions = ['North', 'South', 'East', 'West']

    # pivot_data[product][region] = value or None (blank)
    pivot_data = {
        'Widget':       {'North': 12000, 'South': 8500,  'East': None,  'West': 15000},
        'Gadget':       {'North': None,  'South': 11000, 'East': 9500,  'West': None},
        'Gizmo':        {'North': 14000, 'South': None,  'East': 18000, 'West': 7000},
        'Doohickey':    {'North': 10500, 'South': None,  'East': 13000, 'West': None},
        'Thingamajig':  {'North': None,  'South': 16500, 'East': 9000,  'West': 21000},
    }

    # ========== Sheet 1: Data (raw source) ==========
    ws_data = wb.active
    ws_data.title = 'Data'

    headers_data = ['Product', 'Region', 'Sales']
    for c, h in enumerate(headers_data, 1):
        cell = ws_data.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # Build raw data rows (only for non-None combinations)
    # Split some values into multiple transactions for realism
    raw_rows = [
        ('Widget', 'North', 5000),
        ('Widget', 'North', 7000),
        ('Widget', 'South', 3500),
        ('Widget', 'South', 5000),
        ('Widget', 'West', 8000),
        ('Widget', 'West', 7000),
        ('Gadget', 'South', 6000),
        ('Gadget', 'South', 5000),
        ('Gadget', 'East', 4500),
        ('Gadget', 'East', 5000),
        ('Gizmo', 'North', 9000),
        ('Gizmo', 'North', 5000),
        ('Gizmo', 'East', 10000),
        ('Gizmo', 'East', 8000),
        ('Gizmo', 'West', 7000),
        ('Doohickey', 'North', 4500),
        ('Doohickey', 'North', 6000),
        ('Doohickey', 'East', 8000),
        ('Doohickey', 'East', 5000),
        ('Thingamajig', 'South', 9000),
        ('Thingamajig', 'South', 7500),
        ('Thingamajig', 'East', 4000),
        ('Thingamajig', 'East', 5000),
        ('Thingamajig', 'West', 12000),
        ('Thingamajig', 'West', 9000),
    ]

    for r, (prod, reg, sales) in enumerate(raw_rows, 2):
        ws_data.cell(row=r, column=1, value=prod)
        ws_data.cell(row=r, column=2, value=reg)
        ws_data.cell(row=r, column=3, value=sales)

    # Adjust column widths
    ws_data.column_dimensions['A'].width = 18
    ws_data.column_dimensions['B'].width = 12
    ws_data.column_dimensions['C'].width = 12

    # ========== Sheet 2: PivotSheet (pivot table layout) ==========
    ws_pivot = wb.create_sheet('PivotSheet')

    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    label_font = Font(bold=True)
    center_align = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    number_fmt = '#,##0'

    # Title row
    ws_pivot.merge_cells('A1:F1')
    title_cell = ws_pivot['A1']
    title_cell.value = 'Sum of Sales'
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = Alignment(horizontal="center")

    # Column headers: row 2
    # A2 = "Product", B2 = "North", C2 = "South", D2 = "East", E2 = "West", F2 = "Grand Total"
    col_headers = ['Product'] + regions + ['Grand Total']
    for c, h in enumerate(col_headers, 1):
        cell = ws_pivot.cell(row=2, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Data rows: rows 3-7 (5 products)
    for r, prod in enumerate(products, 3):
        # Product label
        label_cell = ws_pivot.cell(row=r, column=1, value=prod)
        label_cell.font = label_font
        label_cell.border = thin_border

        row_total = 0
        for c, reg in enumerate(regions, 2):
            cell = ws_pivot.cell(row=r, column=c)
            val = pivot_data[prod][reg]
            if val is not None:
                cell.value = val
                cell.number_format = number_fmt
                row_total += val
            # If None, leave cell blank (this is the initial state - no zeros)
            cell.alignment = center_align
            cell.border = thin_border

        # Row grand total
        gt_cell = ws_pivot.cell(row=r, column=6, value=row_total)
        gt_cell.font = Font(bold=True)
        gt_cell.number_format = number_fmt
        gt_cell.alignment = center_align
        gt_cell.border = thin_border

    # Grand Total row (row 8)
    gt_row = 8
    ws_pivot.cell(row=gt_row, column=1, value='Grand Total').font = Font(bold=True)
    ws_pivot.cell(row=gt_row, column=1).border = thin_border
    ws_pivot.cell(row=gt_row, column=1).fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")

    for c, reg in enumerate(regions, 2):
        col_total = sum(pivot_data[p][reg] for p in products if pivot_data[p][reg] is not None)
        cell = ws_pivot.cell(row=gt_row, column=c, value=col_total)
        cell.font = Font(bold=True)
        cell.number_format = number_fmt
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")

    # Grand total of grand totals
    grand_total = 165000
    gt_gt_cell = ws_pivot.cell(row=gt_row, column=6, value=grand_total)
    gt_gt_cell.font = Font(bold=True)
    gt_gt_cell.number_format = number_fmt
    gt_gt_cell.alignment = center_align
    gt_gt_cell.border = thin_border
    gt_gt_cell.fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")

    # Adjust column widths
    ws_pivot.column_dimensions['A'].width = 18
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws_pivot.column_dimensions[col_letter].width = 14

    # Set PivotSheet as active sheet
    wb.active = wb.sheetnames.index('PivotSheet')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
