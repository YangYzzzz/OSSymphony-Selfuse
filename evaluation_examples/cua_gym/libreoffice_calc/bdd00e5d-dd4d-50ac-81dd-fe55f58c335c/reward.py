"""
Reward Script: Show 0 for empty pivot table cells
Task ID: calc_pivot_086
Domain: libreoffice_calc
Scoring:
  Component 1 (0.6): Previously empty cells in pivot table now contain 0
  Component 2 (0.4): Grand total preserved at 165000 AND empty cells fixed
"""

import os

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_086'

# These cells are empty (None) in the initial pivot table but should be 0 in golden.
# PivotSheet layout: Row 2=headers, Rows 3-7=products, Row 8=Grand Total
# Columns: A=Product, B=North, C=South, D=East, E=West, F=Grand Total
EMPTY_CELLS = [
    'D3',  # Widget / East
    'B4',  # Gadget / North
    'E4',  # Gadget / West
    'C5',  # Gizmo / South
    'C6',  # Doohickey / South
    'E6',  # Doohickey / West
    'B7',  # Thingamajig / North
]


def persist_app_state(domain: str):
    """Attempt to save any unsaved GUI state."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: PivotSheet must exist
    if 'PivotSheet' not in wb.sheetnames:
        print("CRITICAL: 'PivotSheet' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['PivotSheet']

    # Component 1: Previously empty cells now contain 0 (0.6 points)
    # Each of the 7 cells contributes proportionally: 0.6/7 per cell
    try:
        zeros_found = 0
        per_cell_score = 0.6 / len(EMPTY_CELLS)
        for coord in EMPTY_CELLS:
            val = ws[coord].value
            if val is not None and (val == 0 or val == 0.0):
                zeros_found += 1
                print(f"PASS: {coord} contains 0 (was empty)")
            else:
                print(f"FAIL: {coord} expected 0, found: {repr(val)}")

        comp1_score = zeros_found * per_cell_score
        if zeros_found > 0:
            print(f"PASS: Component 1 -- {zeros_found}/{len(EMPTY_CELLS)} empty cells now show 0 ({comp1_score:.3f} pts)")
            total_score += comp1_score
        else:
            print(f"FAIL: Component 1 -- no empty cells were changed to 0")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Grand Total preserved at 165000 AND at least one empty cell fixed (0.4 points)
    # This is a compound check: the grand total being 165000 alone is a precondition (true in initial),
    # so we anchor it to the task change by requiring at least one empty cell to also be 0.
    try:
        grand_total = ws['F8'].value
        gt_ok = (grand_total is not None and abs(float(grand_total) - 165000) < 1)

        if gt_ok and zeros_found >= 1:
            print(f"PASS: Component 2 -- Grand Total is {grand_total} and {zeros_found} cells fixed ({0.4} pts)")
            total_score += 0.4
        elif gt_ok and zeros_found == 0:
            print(f"FAIL: Component 2 -- Grand Total is correct ({grand_total}) but no empty cells were fixed")
        else:
            print(f"FAIL: Component 2 -- Grand Total expected 165000, found: {repr(grand_total)}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.3f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
