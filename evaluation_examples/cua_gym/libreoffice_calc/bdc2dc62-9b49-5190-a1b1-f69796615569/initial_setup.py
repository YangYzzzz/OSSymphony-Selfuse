"""
Initial Setup: Insert hyperlink in Contents sheet navigating to Q2 Data sheet
Task ID: calc_ggf_009
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_009'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Contents Sheet (navigation hub) ---
    ws_contents = wb.active
    ws_contents.title = 'Contents'

    # Header
    ws_contents['A1'] = 'Table of Contents'
    ws_contents['A1'].font = Font(name='Arial', size=14, bold=True)
    ws_contents['A1'].alignment = Alignment(horizontal='left')
    ws_contents.column_dimensions['A'].width = 30

    # Subtitle
    ws_contents['A2'] = 'Q1 Data'
    ws_contents['A3'] = 'Q2 Data'
    ws_contents['A4'] = 'Q3 Data'
    ws_contents['A5'] = 'Q4 Data'

    # Plain text styling for all entries (no hyperlink formatting)
    for row in range(2, 6):
        cell = ws_contents.cell(row=row, column=1)
        cell.font = Font(name='Arial', size=11)

    ws_contents['B1'] = 'Description'
    ws_contents['B1'].font = Font(name='Arial', size=14, bold=True)
    ws_contents['B2'] = 'January - March 2025 Sales'
    ws_contents['B3'] = 'April - June 2025 Sales'
    ws_contents['B4'] = 'July - September 2025 Sales'
    ws_contents['B5'] = 'October - December 2025 Sales'
    ws_contents.column_dimensions['B'].width = 35

    # --- Q1 Data Sheet ---
    ws_q1 = wb.create_sheet('Q1 Data')
    q1_headers = ['Product', 'Region', 'January', 'February', 'March', 'Q1 Total']
    for c, h in enumerate(q1_headers, 1):
        cell = ws_q1.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    q1_data = [
        ['Widget Pro', 'North America', 12450, 13200, 14800, 40450],
        ['Widget Pro', 'Europe', 8900, 9100, 10200, 28200],
        ['Widget Pro', 'Asia Pacific', 6700, 7300, 7800, 21800],
        ['Gadget X1', 'North America', 15600, 14900, 16200, 46700],
        ['Gadget X1', 'Europe', 11300, 12100, 11800, 35200],
        ['Gadget X1', 'Asia Pacific', 9200, 9800, 10400, 29400],
        ['Sensor Mark IV', 'North America', 4300, 4700, 5100, 14100],
        ['Sensor Mark IV', 'Europe', 3200, 3500, 3800, 10500],
        ['Sensor Mark IV', 'Asia Pacific', 2800, 3100, 3400, 9300],
        ['CloudSync Module', 'North America', 18200, 19500, 20100, 57800],
        ['CloudSync Module', 'Europe', 14100, 15300, 15800, 45200],
        ['CloudSync Module', 'Asia Pacific', 10600, 11200, 12400, 34200],
    ]
    for r, row_data in enumerate(q1_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_q1.cell(row=r, column=c, value=val)
    ws_q1.column_dimensions['A'].width = 20
    ws_q1.column_dimensions['B'].width = 18

    # --- Q2 Data Sheet ---
    ws_q2 = wb.create_sheet('Q2 Data')
    q2_headers = ['Product', 'Region', 'April', 'May', 'June', 'Q2 Total']
    for c, h in enumerate(q2_headers, 1):
        cell = ws_q2.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='FF548235', end_color='FF548235', fill_type='solid')
    q2_data = [
        ['Widget Pro', 'North America', 15100, 15800, 16200, 47100],
        ['Widget Pro', 'Europe', 10500, 11200, 11800, 33500],
        ['Widget Pro', 'Asia Pacific', 8200, 8900, 9300, 26400],
        ['Gadget X1', 'North America', 16800, 17200, 18100, 52100],
        ['Gadget X1', 'Europe', 12400, 13000, 13500, 38900],
        ['Gadget X1', 'Asia Pacific', 10800, 11300, 12000, 34100],
        ['Sensor Mark IV', 'North America', 5300, 5800, 6200, 17300],
        ['Sensor Mark IV', 'Europe', 4100, 4500, 4800, 13400],
        ['Sensor Mark IV', 'Asia Pacific', 3600, 3900, 4200, 11700],
        ['CloudSync Module', 'North America', 21000, 22300, 23100, 66400],
        ['CloudSync Module', 'Europe', 16200, 17100, 17800, 51100],
        ['CloudSync Module', 'Asia Pacific', 12800, 13500, 14200, 40500],
    ]
    for r, row_data in enumerate(q2_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_q2.cell(row=r, column=c, value=val)
    ws_q2.column_dimensions['A'].width = 20
    ws_q2.column_dimensions['B'].width = 18

    # --- Q3 Data Sheet ---
    ws_q3 = wb.create_sheet('Q3 Data')
    q3_headers = ['Product', 'Region', 'July', 'August', 'September', 'Q3 Total']
    for c, h in enumerate(q3_headers, 1):
        cell = ws_q3.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='FFBF8F00', end_color='FFBF8F00', fill_type='solid')
    q3_data = [
        ['Widget Pro', 'North America', 14200, 13800, 15500, 43500],
        ['Widget Pro', 'Europe', 10100, 9800, 11000, 30900],
        ['Widget Pro', 'Asia Pacific', 7900, 7600, 8500, 24000],
        ['Gadget X1', 'North America', 17500, 16900, 18400, 52800],
        ['Gadget X1', 'Europe', 13100, 12800, 13900, 39800],
        ['Gadget X1', 'Asia Pacific', 11200, 10900, 11800, 33900],
        ['Sensor Mark IV', 'North America', 5800, 5500, 6100, 17400],
        ['Sensor Mark IV', 'Europe', 4400, 4200, 4700, 13300],
        ['Sensor Mark IV', 'Asia Pacific', 3800, 3600, 4100, 11500],
        ['CloudSync Module', 'North America', 22800, 21900, 24100, 68800],
        ['CloudSync Module', 'Europe', 17500, 16800, 18500, 52800],
        ['CloudSync Module', 'Asia Pacific', 13800, 13200, 14600, 41600],
    ]
    for r, row_data in enumerate(q3_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_q3.cell(row=r, column=c, value=val)
    ws_q3.column_dimensions['A'].width = 20
    ws_q3.column_dimensions['B'].width = 18

    # --- Q4 Data Sheet ---
    ws_q4 = wb.create_sheet('Q4 Data')
    q4_headers = ['Product', 'Region', 'October', 'November', 'December', 'Q4 Total']
    for c, h in enumerate(q4_headers, 1):
        cell = ws_q4.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='FFC00000', end_color='FFC00000', fill_type='solid')
    q4_data = [
        ['Widget Pro', 'North America', 16800, 18200, 19500, 54500],
        ['Widget Pro', 'Europe', 12200, 13100, 14000, 39300],
        ['Widget Pro', 'Asia Pacific', 9600, 10300, 11100, 31000],
        ['Gadget X1', 'North America', 19200, 20500, 22100, 61800],
        ['Gadget X1', 'Europe', 14500, 15200, 16100, 45800],
        ['Gadget X1', 'Asia Pacific', 12500, 13100, 14000, 39600],
        ['Sensor Mark IV', 'North America', 6500, 7100, 7800, 21400],
        ['Sensor Mark IV', 'Europe', 5000, 5400, 5900, 16300],
        ['Sensor Mark IV', 'Asia Pacific', 4300, 4700, 5200, 14200],
        ['CloudSync Module', 'North America', 25200, 27100, 29500, 81800],
        ['CloudSync Module', 'Europe', 19500, 20800, 22400, 62700],
        ['CloudSync Module', 'Asia Pacific', 15400, 16500, 17800, 49700],
    ]
    for r, row_data in enumerate(q4_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_q4.cell(row=r, column=c, value=val)
    ws_q4.column_dimensions['A'].width = 20
    ws_q4.column_dimensions['B'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
