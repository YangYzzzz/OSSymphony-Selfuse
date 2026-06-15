"""
Reward Script: Insert hyperlink in cell A3 of Contents sheet to Q2 Data sheet
Task ID: calc_ggf_009
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Cell A3 display text changed to 'Jump to Q2 Data'
  Component 2 (0.30): Cell A3 has a hyperlink object
  Component 3 (0.25): Hyperlink target references 'Q2 Data' sheet cell A1
  Component 4 (0.15): Cell A3 has blue font color and underline (hyperlink styling)
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_009'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Contents' sheet must exist
    if 'Contents' not in wb.sheetnames:
        print("CRITICAL: 'Contents' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Contents']
    cell = ws['A3']

    # Component 1: Cell A3 display text is 'Jump to Q2 Data' (0.30 points)
    # Initial has 'Q2 Data', golden has 'Jump to Q2 Data' — this is a task-introduced change
    try:
        val = cell.value
        if val is not None and str(val).strip() == 'Jump to Q2 Data':
            print(f"PASS: Component 1 — A3 display text is 'Jump to Q2 Data' (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 — expected 'Jump to Q2 Data', found: {repr(val)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Cell A3 has a hyperlink (0.30 points)
    # Initial has no hyperlink, golden has one — this is a task-introduced change
    try:
        hl = cell.hyperlink
        if hl is not None:
            print(f"PASS: Component 2 — A3 has a hyperlink (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 2 — A3 has no hyperlink")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Hyperlink target references 'Q2 Data' sheet at cell A1 (0.25 points)
    # The expected target is "#'Q2 Data'.A1" (internal document link)
    try:
        hl = cell.hyperlink
        if hl is not None:
            target = hl.target or ''
            location = hl.location or ''
            # The hyperlink can be stored as target or location depending on implementation
            combined = (target + ' ' + location).lower()
            # Check that it references Q2 Data sheet and A1 cell
            has_q2_ref = 'q2 data' in combined or 'q2%20data' in combined
            has_a1_ref = 'a1' in combined
            if has_q2_ref and has_a1_ref:
                print(f"PASS: Component 3 — hyperlink targets 'Q2 Data'.A1 (target={repr(target)}, location={repr(location)}) (0.25 pts)")
                total_score += 0.25
            elif has_q2_ref:
                # Partial: references Q2 Data but not specifically A1
                print(f"PARTIAL: Component 3 — hyperlink references Q2 Data but not A1 (target={repr(target)}, location={repr(location)}) (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 3 — hyperlink does not reference Q2 Data (target={repr(target)}, location={repr(location)})")
        else:
            print(f"FAIL: Component 3 — no hyperlink to check target")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Cell A3 has blue font color and underline styling (0.15 points)
    # Initial has no font color and no underline, golden has blue (000000FF) and underline='single'
    try:
        font = cell.font
        has_underline = font.underline is not None and font.underline != 'none'
        has_blue_color = False
        if font.color and font.color.rgb:
            rgb = str(font.color.rgb).upper()
            # Accept various blue shades: 000000FF, FF0000FF, 0000FF, 0563C1, etc.
            # Standard hyperlink blues
            blue_colors = ['000000FF', 'FF0000FF', '0000FF', '0563C1', 'FF0563C1', '000563C1',
                           '0000EE', 'FF0000EE', '1155CC', 'FF1155CC', '4A86C8', 'FF4A86C8']
            if rgb in blue_colors:
                has_blue_color = True
            else:
                # Also check if blue channel is dominant (last 2 chars are high, R and G are low)
                try:
                    # Handle 8-char ARGB or 6-char RGB
                    hex_rgb = rgb[-6:] if len(rgb) >= 6 else rgb
                    r = int(hex_rgb[0:2], 16)
                    g = int(hex_rgb[2:4], 16)
                    b = int(hex_rgb[4:6], 16)
                    if b > 150 and b > r and b > g:
                        has_blue_color = True
                except (ValueError, IndexError):
                    pass

        if has_underline and has_blue_color:
            print(f"PASS: Component 4 — A3 has blue font ({font.color.rgb}) and underline ({font.underline}) (0.15 pts)")
            total_score += 0.15
        elif has_underline or has_blue_color:
            print(f"PARTIAL: Component 4 — A3 has {'underline' if has_underline else 'blue font'} but not {'blue font' if has_underline else 'underline'} (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 4 — A3 has no hyperlink styling (underline={font.underline}, color={font.color.rgb if font.color else None})")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
