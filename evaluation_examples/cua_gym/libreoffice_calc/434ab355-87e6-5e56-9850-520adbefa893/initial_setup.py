"""
Initial Setup: Create a large board meeting report spreadsheet with multiple sections
that span multiple print pages, suitable for page break adjustment practice.
Task ID: calc_gsi_049
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_049'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # Common styles
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    currency_fmt = '$#,##0.00'
    pct_fmt = '0.0%'
    date_fmt = 'yyyy-mm-dd'

    # ========== Sheet 1: Revenue Summary ==========
    ws1 = wb.active
    ws1.title = "Revenue Summary"

    # Title row
    ws1.merge_cells("A1:G1")
    ws1["A1"] = "FY2025 Revenue Summary Report — Board Review"
    ws1["A1"].font = Font(name="Arial", size=14, bold=True, color="1F3864")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    # Subtitle
    ws1.merge_cells("A2:G2")
    ws1["A2"] = "Prepared for Q4 Board Meeting | Confidential"
    ws1["A2"].font = Font(name="Arial", size=10, italic=True, color="666666")
    ws1["A2"].alignment = Alignment(horizontal="center")
    ws1.row_dimensions[2].height = 20

    # Blank row
    ws1.row_dimensions[3].height = 10

    # Headers row 4
    headers1 = ["Region", "Q1 Revenue", "Q2 Revenue", "Q3 Revenue", "Q4 Revenue", "Annual Total", "YoY Growth"]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows — 25 regions for a large dataset
    regions_data = [
        ["North America — East", 4523000, 4891000, 5102000, 5430000, None, 0.082],
        ["North America — West", 3891000, 4102000, 4350000, 4620000, None, 0.065],
        ["North America — Central", 2145000, 2310000, 2480000, 2590000, None, 0.071],
        ["Europe — UK & Ireland", 3210000, 3450000, 3620000, 3890000, None, 0.094],
        ["Europe — DACH Region", 2890000, 3120000, 3280000, 3510000, None, 0.088],
        ["Europe — France & Benelux", 1980000, 2150000, 2310000, 2480000, None, 0.076],
        ["Europe — Southern", 1450000, 1620000, 1780000, 1920000, None, 0.063],
        ["Europe — Nordics", 1230000, 1380000, 1520000, 1650000, None, 0.091],
        ["Asia Pacific — Japan", 2780000, 2950000, 3120000, 3340000, None, 0.102],
        ["Asia Pacific — China", 3450000, 3820000, 4150000, 4520000, None, 0.134],
        ["Asia Pacific — Southeast", 1560000, 1780000, 1950000, 2120000, None, 0.115],
        ["Asia Pacific — ANZ", 1890000, 2050000, 2180000, 2310000, None, 0.078],
        ["Asia Pacific — India", 980000, 1150000, 1340000, 1520000, None, 0.156],
        ["Latin America — Brazil", 1240000, 1380000, 1520000, 1680000, None, 0.092],
        ["Latin America — Mexico", 890000, 980000, 1050000, 1150000, None, 0.073],
        ["Latin America — Other", 560000, 620000, 710000, 790000, None, 0.068],
        ["Middle East — UAE", 780000, 890000, 980000, 1120000, None, 0.118],
        ["Middle East — Saudi Arabia", 650000, 740000, 820000, 950000, None, 0.105],
        ["Africa — South Africa", 420000, 480000, 540000, 610000, None, 0.083],
        ["Africa — Nigeria", 280000, 320000, 370000, 430000, None, 0.097],
        ["Africa — Kenya & East", 180000, 210000, 250000, 290000, None, 0.112],
        ["CIS — Russia", 520000, 480000, 450000, 420000, None, -0.045],
        ["CIS — Other", 340000, 360000, 380000, 410000, None, 0.058],
        ["Caribbean & Islands", 190000, 220000, 260000, 300000, None, 0.089],
        ["Global Strategic Accounts", 5680000, 6120000, 6540000, 7020000, None, 0.121],
    ]

    for r, row_data in enumerate(regions_data, 5):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 1:
                cell.font = Font(name="Arial", size=10)
            elif c == 7:
                cell.number_format = pct_fmt
            elif c >= 2:
                if val is not None:
                    cell.number_format = currency_fmt
        # Annual total formula
        total_cell = ws1.cell(row=r, column=6, value=f"=SUM(B{r}:E{r})")
        total_cell.number_format = currency_fmt
        total_cell.border = thin_border

    last_data_row = 5 + len(regions_data) - 1  # row 29

    # Grand total row
    total_row = last_data_row + 1
    ws1.cell(row=total_row, column=1, value="GRAND TOTAL").font = Font(name="Arial", size=11, bold=True)
    ws1.cell(row=total_row, column=1).border = thin_border
    for col in range(2, 7):
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(col)
        cell = ws1.cell(row=total_row, column=col, value=f"=SUM({col_letter}5:{col_letter}{last_data_row})")
        cell.number_format = currency_fmt
        cell.font = Font(name="Arial", size=11, bold=True)
        cell.border = thin_border
        cell.fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")

    # Set column widths
    ws1.column_dimensions["A"].width = 28
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws1.column_dimensions[col_letter].width = 16
    ws1.column_dimensions["G"].width = 14

    # ========== Sheet 2: Department Budget ==========
    ws2 = wb.create_sheet("Department Budget")

    ws2.merge_cells("A1:H1")
    ws2["A1"] = "Department Budget Allocation & Actuals — FY2025"
    ws2["A1"].font = Font(name="Arial", size=14, bold=True, color="1F3864")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    headers2 = ["Department", "Budget Allocated", "Q1 Actual", "Q2 Actual", "Q3 Actual",
                "Q4 Actual", "Total Spent", "Variance"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    dept_data = [
        ["Engineering", 12500000, 2890000, 3120000, 3250000, 3410000],
        ["Product Management", 4200000, 980000, 1050000, 1120000, 1180000],
        ["Sales & Marketing", 8900000, 2150000, 2340000, 2480000, 2620000],
        ["Customer Success", 3600000, 840000, 910000, 980000, 1050000],
        ["Human Resources", 2800000, 650000, 700000, 740000, 780000],
        ["Finance & Legal", 3100000, 720000, 780000, 810000, 860000],
        ["Operations", 5400000, 1250000, 1380000, 1450000, 1520000],
        ["Research & Innovation", 6200000, 1480000, 1560000, 1640000, 1720000],
        ["IT Infrastructure", 4500000, 1050000, 1120000, 1180000, 1250000],
        ["Data Science & Analytics", 3800000, 890000, 950000, 1020000, 1080000],
        ["Security & Compliance", 2900000, 680000, 720000, 760000, 810000],
        ["Facilities & Admin", 2100000, 490000, 530000, 560000, 590000],
        ["Executive Office", 1800000, 420000, 450000, 470000, 490000],
        ["Training & Development", 1500000, 350000, 380000, 410000, 430000],
        ["Quality Assurance", 2200000, 510000, 550000, 580000, 620000],
        ["Supply Chain", 3400000, 790000, 850000, 900000, 960000],
        ["Partnerships", 1900000, 440000, 480000, 510000, 540000],
        ["Corporate Communications", 1600000, 370000, 400000, 430000, 460000],
        ["Internal Audit", 1100000, 250000, 280000, 300000, 320000],
        ["Business Development", 2700000, 630000, 680000, 720000, 770000],
    ]

    for r, row_data in enumerate(dept_data, 4):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 1:
                cell.font = Font(name="Arial", size=10)
            else:
                cell.number_format = currency_fmt
        # Total Spent
        total_cell = ws2.cell(row=r, column=7, value=f"=SUM(C{r}:F{r})")
        total_cell.number_format = currency_fmt
        total_cell.border = thin_border
        # Variance
        var_cell = ws2.cell(row=r, column=8, value=f"=B{r}-G{r}")
        var_cell.number_format = currency_fmt
        var_cell.border = thin_border

    ws2.column_dimensions["A"].width = 28
    for col_letter in ["B", "C", "D", "E", "F", "G", "H"]:
        ws2.column_dimensions[col_letter].width = 16

    # ========== Sheet 3: Quarterly Targets ==========
    ws3 = wb.create_sheet("Quarterly Targets")

    ws3.merge_cells("A1:I1")
    ws3["A1"] = "Sales Team Quarterly Performance vs Targets"
    ws3["A1"].font = Font(name="Arial", size=14, bold=True, color="1F3864")
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 30

    headers3 = ["Sales Rep", "Territory", "Q1 Target", "Q1 Actual", "Q2 Target",
                "Q2 Actual", "Q3 Target", "Q3 Actual", "Achievement %"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    sales_data = [
        ["Sarah Chen", "Northeast US", 450000, 482000, 480000, 501000, 510000, 535000],
        ["Marcus Johnson", "Southeast US", 380000, 392000, 400000, 415000, 420000, 438000],
        ["Priya Patel", "West Coast", 520000, 498000, 540000, 562000, 570000, 589000],
        ["James O'Brien", "Midwest", 340000, 355000, 360000, 371000, 380000, 394000],
        ["Elena Rodriguez", "Southwest", 290000, 278000, 310000, 325000, 330000, 348000],
        ["Wei Zhang", "Pacific Northwest", 310000, 332000, 330000, 341000, 350000, 368000],
        ["Ahmad Hassan", "UK & Ireland", 420000, 445000, 440000, 458000, 460000, 481000],
        ["Sophie Dubois", "France & Benelux", 360000, 348000, 380000, 392000, 400000, 418000],
        ["Thomas Mueller", "DACH Region", 400000, 421000, 420000, 438000, 440000, 462000],
        ["Maria Rossi", "Southern Europe", 280000, 265000, 300000, 312000, 320000, 338000],
        ["Kenji Tanaka", "Japan", 480000, 502000, 500000, 518000, 520000, 545000],
        ["Li Ming", "Greater China", 550000, 589000, 580000, 612000, 610000, 648000],
        ["Rajesh Kumar", "India & SE Asia", 320000, 341000, 340000, 358000, 360000, 382000],
        ["Carlos Silva", "Brazil", 260000, 248000, 280000, 295000, 300000, 318000],
        ["Fatima Al-Rashid", "Middle East", 350000, 372000, 370000, 388000, 390000, 412000],
        ["David Okafor", "Africa", 220000, 208000, 240000, 252000, 260000, 278000],
        ["Anna Kowalski", "Eastern Europe", 300000, 312000, 320000, 335000, 340000, 358000],
        ["Ryan Thompson", "Canada", 370000, 385000, 390000, 401000, 410000, 428000],
        ["Isabella Santos", "Latin America Other", 240000, 228000, 260000, 275000, 280000, 298000],
        ["Henrik Johansson", "Nordics", 330000, 348000, 350000, 365000, 370000, 392000],
    ]

    for r, row_data in enumerate(sales_data, 4):
        for c, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c <= 2:
                cell.font = Font(name="Arial", size=10)
            else:
                cell.number_format = currency_fmt
        # Achievement % formula
        ach_cell = ws3.cell(row=r, column=9, value=f"=(D{r}+F{r}+H{r})/(C{r}+E{r}+G{r})")
        ach_cell.number_format = pct_fmt
        ach_cell.border = thin_border

    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 22
    for col_letter in ["C", "D", "E", "F", "G", "H"]:
        ws3.column_dimensions[col_letter].width = 15
    ws3.column_dimensions["I"].width = 16

    # ========== Sheet 4: Monthly Breakdown ==========
    ws4 = wb.create_sheet("Monthly Breakdown")

    ws4.merge_cells("A1:N1")
    ws4["A1"] = "Monthly Revenue Breakdown by Product Line"
    ws4["A1"].font = Font(name="Arial", size=14, bold=True, color="1F3864")
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 30

    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    headers4 = ["Product Line"] + months + ["Annual Total"]
    for col, h in enumerate(headers4, 1):
        cell = ws4.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    products = [
        ["Enterprise SaaS Platform", 890000, 920000, 950000, 980000, 1020000, 1050000, 1080000, 1120000, 1150000, 1190000, 1230000, 1280000],
        ["SMB Cloud Suite", 450000, 465000, 480000, 498000, 515000, 530000, 548000, 565000, 582000, 600000, 618000, 640000],
        ["Data Analytics Pro", 320000, 335000, 348000, 362000, 378000, 392000, 408000, 422000, 438000, 455000, 470000, 488000],
        ["Security Shield", 280000, 290000, 302000, 315000, 328000, 342000, 355000, 370000, 385000, 398000, 412000, 428000],
        ["DevOps Toolkit", 210000, 218000, 228000, 238000, 248000, 258000, 270000, 280000, 292000, 305000, 318000, 332000],
        ["Mobile Workspace", 180000, 188000, 195000, 204000, 212000, 222000, 232000, 242000, 252000, 264000, 275000, 288000],
        ["AI Assistant Module", 120000, 135000, 152000, 170000, 190000, 212000, 238000, 265000, 295000, 328000, 365000, 408000],
        ["Compliance Manager", 150000, 155000, 162000, 168000, 175000, 182000, 190000, 198000, 206000, 215000, 224000, 234000],
        ["Integration Hub", 95000, 100000, 105000, 112000, 118000, 125000, 132000, 140000, 148000, 156000, 165000, 175000],
        ["Customer Portal", 140000, 148000, 155000, 162000, 170000, 178000, 186000, 195000, 204000, 214000, 224000, 235000],
        ["Reporting Engine", 110000, 115000, 120000, 126000, 132000, 138000, 145000, 152000, 160000, 168000, 176000, 185000],
        ["Workflow Automation", 175000, 182000, 190000, 198000, 208000, 218000, 228000, 240000, 252000, 264000, 278000, 292000],
        ["API Gateway", 85000, 90000, 95000, 100000, 108000, 115000, 122000, 130000, 138000, 148000, 158000, 168000],
        ["Document Management", 130000, 135000, 142000, 148000, 155000, 162000, 170000, 178000, 186000, 195000, 205000, 215000],
        ["Training Platform", 92000, 96000, 100000, 106000, 112000, 118000, 125000, 132000, 140000, 148000, 158000, 168000],
    ]

    for r, row_data in enumerate(products, 4):
        for c, val in enumerate(row_data, 1):
            cell = ws4.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 1:
                cell.font = Font(name="Arial", size=10)
            else:
                cell.number_format = currency_fmt
        # Annual Total formula
        total_cell = ws4.cell(row=r, column=14, value=f"=SUM(B{r}:M{r})")
        total_cell.number_format = currency_fmt
        total_cell.border = thin_border
        total_cell.font = Font(name="Arial", size=10, bold=True)

    ws4.column_dimensions["A"].width = 26
    for col_letter in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]:
        ws4.column_dimensions[col_letter].width = 13

    # Save workbook
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
