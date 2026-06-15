"""
Reward Script: Delete all comments across the entire sheet in a single operation
Task ID: calc_cop_comment_006
Domain: libreoffice_calc
Scoring:
  - Component 1: Total comment count is zero (0.6 pts) — fails on initial (15 comments), passes on golden (0)
  - Component 2: All 15 originally-commented cells are individually comment-free (0.4 pts)
    — validates completeness: every specific cell that had a comment has been cleaned
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_cop_comment_006'
SHEET_NAME = 'AuditReady'

# The 15 cells that had comments in the initial file.
# Component 2 verifies each of these is now comment-free in the golden/submitted file.
ORIGINALLY_COMMENTED_CELLS = [
    'B3', 'E6', 'G6', 'C10', 'H10',
    'A13', 'E13', 'G17', 'B18', 'E18',
    'H23', 'C28', 'B32', 'E34', 'H39'
]


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet 'AuditReady' must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Component 1: Total comment count in the sheet is zero (0.6 points)
    # The initial file has 15 comments; after task completion there should be 0.
    # This FAILS on initial (15 comments exist) and PASSES on golden (0 comments).
    try:
        all_comments = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment is not None:
                    all_comments.append(cell.coordinate)

        comment_count = len(all_comments)
        if comment_count == 0:
            print(f"PASS: Component 1 — Sheet '{SHEET_NAME}' has 0 comments (all removed) (0.6 pts)")
            total_score += 0.6
        else:
            print(f"FAIL: Component 1 — Expected 0 comments, found {comment_count} comment(s) at: {all_comments[:10]}")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not count comments: {e}")

    # Component 2: Each of the 15 originally-commented cells is individually comment-free (0.4 points)
    # This verifies completeness — not just that the count is zero, but that every specific
    # cell that previously had a comment has been cleaned.
    # This FAILS on initial (all 15 cells have comments) and PASSES on golden (all are clean).
    try:
        cells_still_with_comments = []
        for coord in ORIGINALLY_COMMENTED_CELLS:
            cell = ws[coord]
            if cell.comment is not None:
                cells_still_with_comments.append(coord)

        if len(cells_still_with_comments) == 0:
            print(f"PASS: Component 2 — All {len(ORIGINALLY_COMMENTED_CELLS)} originally-commented cells are now comment-free (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 2 — {len(cells_still_with_comments)} of {len(ORIGINALLY_COMMENTED_CELLS)} originally-commented cells still have comments: {cells_still_with_comments}")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not check individual cells: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
