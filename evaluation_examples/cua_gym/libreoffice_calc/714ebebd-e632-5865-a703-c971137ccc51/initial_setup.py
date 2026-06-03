"""
Initial Setup: Spreadsheet with comments scattered across AuditReady sheet
Task ID: calc_cop_comment_006
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

WORKDIR = '/home/user'
TASK_ID = 'calc_cop_comment_006'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: AuditReady ---
    ws = wb.active
    ws.title = 'AuditReady'

    # Styles
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF17375E', end_color='FF17375E', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='000000')
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    alt_fill_blue = PatternFill(start_color='FFDCE6F1', end_color='FFDCE6F1', fill_type='solid')
    alt_fill_white = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')

    flag_fill_red = PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')
    flag_fill_green = PatternFill(start_color='FFC6EFCE', end_color='FFC6EFCE', fill_type='solid')

    # Column headers: 8 columns
    headers = [
        'Vendor ID', 'Vendor Name', 'Invoice #', 'Invoice Date',
        'Amount (USD)', 'Department', 'Status', 'Auditor Notes'
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border_all

    ws.row_dimensions[1].height = 28
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 30

    # 40 rows of realistic vendor/audit data
    data = [
        ('V-1001', 'Apex Solutions LLC',        'INV-20501', '2025-01-05',  12450.00, 'Engineering',  'Approved',  'Cleared by Q1 audit'),
        ('V-1002', 'Meridian Tech Partners',    'INV-20502', '2025-01-08',   8320.75, 'IT',           'Approved',  ''),
        ('V-1003', 'BlueStar Consulting',       'INV-20503', '2025-01-10',  33100.00, 'Operations',   'Under Review', 'Awaiting CFO sign-off'),
        ('V-1004', 'Harbor Freight Services',   'INV-20504', '2025-01-12',   5670.50, 'Facilities',   'Approved',  ''),
        ('V-1005', 'Granite Peak Industries',   'INV-20505', '2025-01-15',  19800.00, 'Manufacturing','Approved',  ''),
        ('V-1006', 'Cascade Data Systems',      'INV-20506', '2025-01-17',  47500.00, 'IT',           'Flagged',   'Duplicate payment risk'),
        ('V-1007', 'Ironwood Capital Group',    'INV-20507', '2025-01-20',   3250.00, 'Finance',      'Approved',  ''),
        ('V-1008', 'Coastal Print & Media',     'INV-20508', '2025-01-22',   1890.25, 'Marketing',    'Approved',  ''),
        ('V-1009', 'Delta Logistics Corp',      'INV-20509', '2025-01-24',  62400.00, 'Operations',   'Approved',  ''),
        ('V-1010', 'Forrest Analytics',         'INV-20510', '2025-01-27',  15300.00, 'Analytics',    'Under Review', 'Rate variance noted'),
        ('V-1011', 'Summit Office Supplies',    'INV-20511', '2025-01-29',   2140.00, 'Admin',        'Approved',  ''),
        ('V-1012', 'NovaTech Engineering',      'INV-20512', '2025-02-02',  88750.00, 'Engineering',  'Approved',  ''),
        ('V-1013', 'Pacific Rim Exporters',     'INV-20513', '2025-02-04',  21600.00, 'Procurement',  'Flagged',   'Exceeds PO threshold'),
        ('V-1014', 'Sunrise HR Solutions',      'INV-20514', '2025-02-06',   9450.00, 'HR',           'Approved',  ''),
        ('V-1015', 'TerraGreen Landscaping',    'INV-20515', '2025-02-09',   4780.00, 'Facilities',   'Approved',  ''),
        ('V-1016', 'Monarch Software Inc',      'INV-20516', '2025-02-11',  31200.00, 'IT',           'Approved',  ''),
        ('V-1017', 'Bridgewater Consulting',    'INV-20517', '2025-02-13',  17800.00, 'Strategy',     'Under Review', 'Contract renewal pending'),
        ('V-1018', 'Apex Solutions LLC',        'INV-20518', '2025-02-15',  12450.00, 'Engineering',  'Flagged',   'Possible duplicate of INV-20501'),
        ('V-1019', 'Redwood Staffing Group',    'INV-20519', '2025-02-17',  43600.00, 'HR',           'Approved',  ''),
        ('V-1020', 'Clearwater Chemicals',      'INV-20520', '2025-02-19',  76200.00, 'Manufacturing','Approved',  ''),
        ('V-1021', 'Northern Lights Travel',    'INV-20521', '2025-02-21',   5990.00, 'Admin',        'Approved',  ''),
        ('V-1022', 'Pinnacle Security Co',      'INV-20522', '2025-02-24',  11350.00, 'Facilities',   'Approved',  ''),
        ('V-1023', 'Velocity Marketing',        'INV-20523', '2025-02-26',  28400.00, 'Marketing',    'Flagged',   'Unclear deliverables'),
        ('V-1024', 'Eagle Eye Analytics',       'INV-20524', '2025-02-27',   7830.00, 'Analytics',    'Approved',  ''),
        ('V-1025', 'Westport Fleet Management', 'INV-20525', '2025-03-01',  14200.00, 'Operations',   'Approved',  ''),
        ('V-1026', 'Sterling Payroll Services', 'INV-20526', '2025-03-03',  38900.00, 'Finance',      'Approved',  ''),
        ('V-1027', 'Crescent Moon Catering',    'INV-20527', '2025-03-05',   6430.00, 'Admin',        'Approved',  ''),
        ('V-1028', 'Atlas Cloud Services',      'INV-20528', '2025-03-07',  95400.00, 'IT',           'Under Review', 'SLA compliance check'),
        ('V-1029', 'Glacier Water Systems',     'INV-20529', '2025-03-09',   3120.00, 'Facilities',   'Approved',  ''),
        ('V-1030', 'Titan Printing Solutions',  'INV-20530', '2025-03-11',   8760.00, 'Marketing',    'Approved',  ''),
        ('V-1031', 'Vanguard Legal Services',   'INV-20531', '2025-03-13',  52000.00, 'Legal',        'Approved',  ''),
        ('V-1032', 'Silver Oak Research',       'INV-20532', '2025-03-15',  18300.00, 'R&D',          'Under Review', 'IP ownership clause'),
        ('V-1033', 'Horizon Event Planning',    'INV-20533', '2025-03-17',   7150.00, 'Admin',        'Approved',  ''),
        ('V-1034', 'Oakdale Construction',      'INV-20534', '2025-03-19',  143000.00,'Facilities',   'Flagged',   'Over budget approval needed'),
        ('V-1035', 'Prime Recruiting Partners', 'INV-20535', '2025-03-21',  22700.00, 'HR',           'Approved',  ''),
        ('V-1036', 'Nimbus Cloud Backup',       'INV-20536', '2025-03-23',  11200.00, 'IT',           'Approved',  ''),
        ('V-1037', 'Coastal Freight Express',   'INV-20537', '2025-03-25',  29500.00, 'Operations',   'Approved',  ''),
        ('V-1038', 'Matrix Training Group',     'INV-20538', '2025-03-27',  15600.00, 'HR',           'Approved',  ''),
        ('V-1039', 'Quantum Sensor Labs',       'INV-20539', '2025-03-29',  67800.00, 'R&D',          'Under Review', 'Export control review'),
        ('V-1040', 'Zenith Office Interiors',   'INV-20540', '2025-03-31',  34500.00, 'Facilities',   'Approved',  ''),
    ]

    status_colors = {
        'Approved':      flag_fill_green,
        'Under Review':  alt_fill_blue,
        'Flagged':       flag_fill_red,
    }

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border_all
            # Alternating row background for non-status rows
            if c != 7:
                if r % 2 == 0:
                    cell.fill = alt_fill_white
                else:
                    cell.fill = alt_fill_blue
            else:
                # Status column gets color-coded
                cell.fill = status_colors.get(str(val), alt_fill_white)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            # Format amount column
            if c == 5:
                cell.number_format = '#,##0.00'
            # Format date column
            if c == 4:
                cell.number_format = 'yyyy-mm-dd'

    # Bold the vendor name for Flagged rows
    flagged_rows = [6, 13, 18, 23, 34]  # 1-indexed data rows in data list → actual rows = +1 (header)
    for r_idx in flagged_rows:
        ws.cell(row=r_idx + 1, column=2).font = Font(name='Calibri', size=11, bold=True)

    # Freeze header row
    ws.freeze_panes = 'A2'

    # --- Add 15 comments scattered throughout the sheet ---
    # Comments with realistic audit reviewer content
    comment_cells = [
        ('B3',  'Elena Marsh',  'Verify SOW attachment before final approval.'),
        ('E6',  'James Holt',   'Flag: Cross-check with AP system for duplicate entry.'),
        ('G6',  'James Holt',   'Hold payment until investigation complete.'),
        ('C10', 'Priya Nair',   'Rate differs from agreed contract rate in exhibit B.'),
        ('H10', 'Priya Nair',   'Need written explanation from department head.'),
        ('A13', 'Tom Garcia',   'PO limit is $20,000. Escalation required.'),
        ('E13', 'Tom Garcia',   'Amount exceeds authorized PO by 8%. Requires VP approval.'),
        ('G17', 'Lisa Fernandez', 'Contract expired Feb 28. Renew before processing.'),
        ('B18', 'Elena Marsh',  'Possible duplicate of V-1001 invoice from January.'),
        ('E18', 'Elena Marsh',  'Do NOT pay until reconciliation with INV-20501 is complete.'),
        ('H23', 'James Holt',   'Request deliverable list and sign-off documentation.'),
        ('C28', 'Priya Nair',   'SLA metrics report missing. On hold pending submission.'),
        ('B32', 'Lisa Fernandez', 'IP agreement clause 4.3 may conflict with company policy.'),
        ('E34', 'Tom Garcia',   'Budget overrun by $43k. Board approval required per policy 7.2.'),
        ('H39', 'Priya Nair',   'ECCN classification pending — legal review in progress.'),
    ]

    for coord, author, text in comment_cells:
        c = Comment(text, author)
        c.width = 250
        c.height = 80
        ws[coord].comment = c

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet: AuditReady, 40 rows x 8 cols')
    print(f'Comments added to: {[coord for coord, _, _ in comment_cells]}')


create_initial()
