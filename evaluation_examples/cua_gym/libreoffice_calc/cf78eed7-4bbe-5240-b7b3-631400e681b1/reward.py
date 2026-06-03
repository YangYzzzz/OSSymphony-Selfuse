"""
Reward Script: Build a shipping manifest template for outbound orders
Task ID: calc_ops_warehouse_shipping_manifest_019
Domain: libreoffice_calc
Scoring:
  Component 1: Merged header section (A1:H3, A4:D6, E4:H6)   — 0.25 pts
  Component 2: Header cell content and styling (A1 bold, large, centered)  — 0.20 pts
  Component 3: Row 8 column headers (all 8 labels present)    — 0.15 pts
  Component 4: Calculation formulas in F9:F23 and H9:H23      — 0.20 pts
  Component 5: Totals row 24 (SUM formulas in F24 and H24)    — 0.10 pts
  Component 6: Print settings (area, landscape, fit-to-page)  — 0.10 pts
  Total: 1.0
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_ops_warehouse_shipping_manifest_019'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that the ShipManifest sheet exists (precondition gate)
    if 'ShipManifest' not in wb.sheetnames:
        print("FAIL: 'ShipManifest' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['ShipManifest']

    # -----------------------------------------------------------------------
    # Component 1: Merged cell ranges (0.25 points)
    # Three merged regions must exist: A1:H3, A4:D6, E4:H6
    # -----------------------------------------------------------------------
    try:
        merge_ranges = {str(mr) for mr in ws.merged_cells.ranges}

        has_header_merge = 'A1:H3' in merge_ranges
        has_shipper_merge = 'A4:D6' in merge_ranges
        has_consignee_merge = 'E4:H6' in merge_ranges

        if has_header_merge and has_shipper_merge and has_consignee_merge:
            print("PASS: Component 1 — All three merge regions present (A1:H3, A4:D6, E4:H6) (0.25 pts)")
            total_score += 0.25
        elif has_header_merge and (has_shipper_merge or has_consignee_merge):
            print(f"PARTIAL: Component 1 — Header merge present but shipper={has_shipper_merge}, consignee={has_consignee_merge} (0.15 pts)")
            total_score += 0.15
        elif has_header_merge:
            print("PARTIAL: Component 1 — Header merge A1:H3 present but missing shipper/consignee merges (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 1 — Expected merges A1:H3, A4:D6, E4:H6; found: {merge_ranges}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Header cell content and styling (0.20 points)
    # A1 should contain 'Acme Logistics Ltd', bold, large font, centered
    # -----------------------------------------------------------------------
    try:
        a1 = ws['A1']
        a1_value = a1.value

        has_company_name = a1_value is not None and 'Acme Logistics Ltd' in str(a1_value)
        has_bold = a1.font.bold == True
        has_large_font = a1.font.size is not None and a1.font.size >= 14
        has_centered = a1.alignment.horizontal in ('center', 'centerContinuous')

        checks_passed = sum([has_company_name, has_bold, has_large_font, has_centered])

        if checks_passed == 4:
            print(f"PASS: Component 2 — A1 has company name '{a1_value}', bold={has_bold}, size={a1.font.size}, centered (0.20 pts)")
            total_score += 0.20
        elif checks_passed >= 2 and has_company_name:
            print(f"PARTIAL: Component 2 — A1 has company name but styling partial: bold={has_bold}, size={a1.font.size}, centered={has_centered} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 — A1 value={repr(a1_value)}, bold={has_bold}, size={a1.font.size}, centered={has_centered}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Row 8 column headers (0.15 points)
    # Expected: Line #, SKU, Description, Qty, Unit Weight kg, Total Weight, Unit Value, Total Value
    # -----------------------------------------------------------------------
    try:
        expected_headers = ['Line #', 'SKU', 'Description', 'Qty', 'Unit Weight kg', 'Total Weight', 'Unit Value', 'Total Value']
        actual_headers = [ws.cell(row=8, column=col).value for col in range(1, 9)]

        matching = sum(1 for exp, act in zip(expected_headers, actual_headers)
                       if act is not None and str(act).strip() == exp)

        if matching == 8:
            print(f"PASS: Component 3 — All 8 row-8 headers present (0.15 pts)")
            total_score += 0.15
        elif matching >= 5:
            print(f"PARTIAL: Component 3 — {matching}/8 headers match in row 8 (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 3 — Only {matching}/8 headers match; actual={actual_headers}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: Calculation formulas F9:F23 (=D*E) and H9:H23 (=D*G) (0.20 points)
    # Each row 9-23 should have F=D*E and H=D*G
    # -----------------------------------------------------------------------
    try:
        f_formula_count = 0
        h_formula_count = 0

        for row in range(9, 24):
            f_cell = ws.cell(row=row, column=6)  # Column F
            h_cell = ws.cell(row=row, column=8)  # Column H

            # Check F formula: =D{row}*E{row}
            if f_cell.value and isinstance(f_cell.value, str):
                normalized = f_cell.value.upper().replace(' ', '').replace('=', '')
                expected_f = f'D{row}*E{row}'
                if normalized == expected_f:
                    f_formula_count += 1

            # Check H formula: =D{row}*G{row}
            if h_cell.value and isinstance(h_cell.value, str):
                normalized = h_cell.value.upper().replace(' ', '').replace('=', '')
                expected_h = f'D{row}*G{row}'
                if normalized == expected_h:
                    h_formula_count += 1

        total_formula_checks = f_formula_count + h_formula_count
        max_possible = 30  # 15 rows * 2 formula columns

        if total_formula_checks == max_possible:
            print(f"PASS: Component 4 — All 30 calculation formulas present (F9:F23=D*E, H9:H23=D*G) (0.20 pts)")
            total_score += 0.20
        elif total_formula_checks >= 20:
            print(f"PARTIAL: Component 4 — {total_formula_checks}/{max_possible} formulas present (0.12 pts)")
            total_score += 0.12
        elif total_formula_checks >= 10:
            print(f"PARTIAL: Component 4 — {total_formula_checks}/{max_possible} formulas present (0.06 pts)")
            total_score += 0.06
        else:
            print(f"FAIL: Component 4 — Only {total_formula_checks}/{max_possible} formulas (F={f_formula_count}/15, H={h_formula_count}/15)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -----------------------------------------------------------------------
    # Component 5: Row 24 totals row with SUM formulas (0.10 points)
    # F24 = =SUM(F9:F23), H24 = =SUM(H9:H23)
    # -----------------------------------------------------------------------
    try:
        f24 = ws.cell(row=24, column=6).value
        h24 = ws.cell(row=24, column=8).value

        f24_ok = (f24 is not None and isinstance(f24, str) and
                  'SUM' in f24.upper() and 'F9' in f24.upper() and 'F23' in f24.upper())
        h24_ok = (h24 is not None and isinstance(h24, str) and
                  'SUM' in h24.upper() and 'H9' in h24.upper() and 'H23' in h24.upper())

        if f24_ok and h24_ok:
            print(f"PASS: Component 5 — F24={repr(f24)}, H24={repr(h24)} (0.10 pts)")
            total_score += 0.10
        elif f24_ok or h24_ok:
            print(f"PARTIAL: Component 5 — F24={repr(f24)}, H24={repr(h24)} — only one SUM formula correct (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — F24={repr(f24)}, H24={repr(h24)}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # -----------------------------------------------------------------------
    # Component 6: Print settings — area A1:H24, landscape, fit to 1 page (0.10 points)
    # -----------------------------------------------------------------------
    try:
        print_area = ws.print_area or ''
        is_landscape = ws.page_setup.orientation == 'landscape'

        # Check fit to page setting
        fit_to_page = False
        try:
            fit_to_page = ws.sheet_properties.pageSetUpPr.fitToPage == True
        except (AttributeError, TypeError):
            # Alternatively check fitToWidth and fitToHeight
            try:
                fit_to_page = (ws.page_setup.fitToWidth == 1 and ws.page_setup.fitToHeight == 1)
            except Exception:
                fit_to_page = False

        # Print area check — normalize to remove absolute refs ($)
        normalized_print_area = print_area.replace('$', '').replace("'ShipManifest'!", '')
        has_print_area = 'A1:H24' in normalized_print_area

        checks = [has_print_area, is_landscape, fit_to_page]
        checks_passed = sum(checks)

        if checks_passed == 3:
            print(f"PASS: Component 6 — Print area={print_area}, landscape={is_landscape}, fit_to_page={fit_to_page} (0.10 pts)")
            total_score += 0.10
        elif checks_passed == 2:
            print(f"PARTIAL: Component 6 — 2/3 print settings correct: area={has_print_area}, landscape={is_landscape}, fit={fit_to_page} (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 — print_area={print_area}, landscape={is_landscape}, fit_to_page={fit_to_page}")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
