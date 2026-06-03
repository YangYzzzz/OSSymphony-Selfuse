"""
Reward Script: Read statistical_analysis_plan.docx and execute all statistical analysis steps in survey_results.xlsx
Task ID: osworld_multi_apps_docx_to_calc_015
Domain: libreoffice_calc
Scoring:
  Component 1: 'Analysis' sheet exists in survey_results.xlsx (0.20 pts)
  Component 2: Section 1 - Descriptive statistics (AVERAGE, STDEV, MIN, MAX) formulas present for Q1-Q5 (0.25 pts)
  Component 3: Section 2 - COUNTIF frequency distribution for ratings 1-5 across Q1-Q5 (0.20 pts)
  Component 4: Section 3 - AVERAGEIF by Department for Q1-Q5 (0.20 pts)
  Component 5: Section 4 - NPS calculation (promoters, detractors, NPS score) (0.15 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_docx_to_calc_015'


def cell_has_formula(ws, row_num, col_range_start, col_range_end, keyword):
    """
    Return True if any cell in the given row and column range contains
    a formula string that includes the specified keyword (case-insensitive).
    This function performs actual API reads to verify formula presence.
    """
    for col_num in range(col_range_start, col_range_end + 1):
        cell_val = ws.cell(row=row_num, column=col_num).value
        if isinstance(cell_val, str) and keyword.upper() in cell_val.upper():
            return True
    return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    The task requires creating an 'Analysis' sheet in survey_results.xlsx
    with all statistical measures derived from survey data as described in the docx.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: 'Analysis' sheet exists (0.20 points)
    # This is the primary structural change introduced by the task —
    # the initial file only has 'Survey Data', the golden adds 'Analysis'.
    try:
        if 'Analysis' in wb.sheetnames:
            print("PASS: Component 1 — 'Analysis' sheet exists (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 — 'Analysis' sheet not found. Sheets: {wb.sheetnames}")
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws = wb['Analysis']

    # Component 2: Section 1 - Descriptive statistics (AVERAGE, STDEV, MIN, MAX) (0.25 points)
    # Rows 3-6 should have AVERAGE, STDEV, MIN, MAX formulas for Q1-Q5 (columns B-F)
    # Each formula type must appear in its designated row across the 5 question columns.
    try:
        # formula_types: (row_num, expected_function_name)
        formula_checks = [
            (3, 'AVERAGE'),
            (4, 'STDEV'),
            (5, 'MIN'),
            (6, 'MAX'),
        ]
        stats_results = {
            fname: cell_has_formula(ws, rnum, 2, 6, fname)
            for rnum, fname in formula_checks
        }
        missing_stats = [fname for fname, found in stats_results.items() if not found]

        if len(missing_stats) == 0:
            print("PASS: Component 2 — All 4 descriptive stat functions (AVERAGE/STDEV/MIN/MAX) found in Analysis rows 3-6 (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — Missing descriptive statistics in Analysis rows 3-6: {missing_stats}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Section 2 - COUNTIF frequency distribution for ratings 1-5 (0.20 points)
    # Rows 10-14 should have COUNTIF formulas for ratings 1-5 across Q1-Q5 (columns B-F).
    try:
        countif_row_count = sum(
            1 for row_num in range(10, 15)
            if cell_has_formula(ws, row_num, 2, 6, 'COUNTIF')
        )

        if countif_row_count >= 5:
            print(f"PASS: Component 3 — COUNTIF frequency distribution found for all 5 ratings (0.20 pts)")
            total_score += 0.20
        elif countif_row_count >= 3:
            print(f"PARTIAL: Component 3 — COUNTIF found for {countif_row_count}/5 ratings, partial credit (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — COUNTIF frequency distribution insufficient: only {countif_row_count}/5 rating rows have formulas")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Section 3 - AVERAGEIF by Department (0.20 points)
    # Rows 17+ should have AVERAGEIF formulas per department (at least 6 departments expected).
    try:
        averageif_row_count = sum(
            1 for row_num in range(17, 30)
            if cell_has_formula(ws, row_num, 2, 6, 'AVERAGEIF')
        )

        if averageif_row_count >= 6:
            print(f"PASS: Component 4 — AVERAGEIF department analysis found for {averageif_row_count} departments (0.20 pts)")
            total_score += 0.20
        elif averageif_row_count >= 3:
            print(f"PARTIAL: Component 4 — AVERAGEIF found for {averageif_row_count}/7+ departments, partial credit (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 — AVERAGEIF by department insufficient: only {averageif_row_count} rows found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Section 4 - NPS calculation (0.15 points)
    # NPS section (rows 27+) should contain:
    #   - Promoters count via COUNTIF for NPS values 9-10
    #   - Detractors count via COUNTIF for NPS values 0-6 (e.g., "<=6")
    #   - NPS score formula that computes (Promoters% - Detractors%) * 100
    try:
        nps_rows_values = [
            ws.cell(row=row_num, column=col_num).value
            for row_num in range(27, 40)
            for col_num in range(1, 4)
        ]
        nps_strings = [v for v in nps_rows_values if isinstance(v, str)]

        promoter_count = sum(
            1 for v in nps_strings
            if 'COUNTIF' in v.upper() and ('9' in v or '10' in v)
        )
        detractor_count = sum(
            1 for v in nps_strings
            if 'COUNTIF' in v.upper() and ('<=6' in v or '<= 6' in v)
        )
        nps_formula_count = sum(
            1 for v in nps_strings
            if '-' in v and '100' in v
        )

        nps_components_found = (
            (1 if promoter_count >= 1 else 0) +
            (1 if detractor_count >= 1 else 0) +
            (1 if nps_formula_count >= 1 else 0)
        )

        if nps_components_found >= 3:
            print(f"PASS: Component 5 — NPS calculation complete: promoters, detractors, and NPS formula found (0.15 pts)")
            total_score += 0.15
        elif nps_components_found >= 2:
            print(f"PARTIAL: Component 5 — Partial NPS: {nps_components_found}/3 NPS components found (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 5 — NPS calculation insufficient: {nps_components_found}/3 components found "
                  f"(promoters={promoter_count >= 1}, detractors={detractor_count >= 1}, nps_formula={nps_formula_count >= 1})")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against canonical artifact path in the VM env
file_path = f'{WORKDIR}/survey_results.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
