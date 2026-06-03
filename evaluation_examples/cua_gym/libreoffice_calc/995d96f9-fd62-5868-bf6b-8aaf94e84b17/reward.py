"""
Reward Script: Extract PDF paper info into a sorted Calc spreadsheet
Task ID: osworld_multi_apps_pdf_author_extract_011
Domain: libreoffice_calc (multi-app: PDF + Calc)
Scoring:
  Component 1: File exists with correct headers and 7 data rows (0.3 pts)
  Component 2: All 7 rows have correct Author Name, Email, Benchmark Name, Year (0.5 pts)
  Component 3: Data sorted by Year ascending, then by Author Name alphabetically (0.2 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_pdf_author_extract_011'
FILE_PATH = f'{WORKDIR}/benchmark_authors.xlsx'

# Expected data (sorted by Year ascending, then Author Name alphabetically)
EXPECTED_DATA = [
    ('Olga Russakovsky',  'olga@cs.princeton.edu',          'ImageNet Large Scale Visual Recognition Challenge',                                                       2015),
    ('Pranav Rajpurkar',  'pranavsr@cs.stanford.edu',       'SQuAD: 100,000+ Questions for Machine Comprehension of Text',                                             2016),
    ('Alex Wang',         'alexwang@nyu.edu',               'GLUE: A Multi-Task Benchmark and Analysis Platform for Natural Language Understanding',                    2018),
    ('Rowan Zellers',     'rzellers@cs.washington.edu',     'HellaSwag: Can a Machine Really Finish Your Sentence?',                                                    2019),
    ('Dan Hendrycks',     'dan@berkeley.edu',               'Measuring Massive Multitask Language Understanding',                                                       2021),
    ('Mark Chen',         'mark@openai.com',                'Evaluating Large Language Models Trained on Code (HumanEval)',                                             2021),
    ('Aarohi Srivastava', 'aarohi@google.com',              'Beyond the Imitation Game: Quantifying and Extrapolating the Capabilities of Language Models',            2022),
]

EXPECTED_HEADERS = ['Author Name', 'Email', 'Benchmark Name', 'Year']


def normalize(val):
    """Normalize a cell value to string, stripping whitespace."""
    if val is None:
        return ''
    return str(val).strip()


def year_to_int(val):
    """Convert year cell value to int for comparison."""
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition gate: File must exist
    if not os.path.exists(file_path):
        print(f"CRITICAL: Output file not found: {file_path}")
        print("REWARD: 0.0")
        return 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load workbook {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Use the active/first sheet
    ws = wb.active

    # -------------------------------------------------------------------------
    # Component 1: Correct headers and 7 data rows (0.3 points)
    # This FAILS on initial_env (file does not exist) and PASSES on golden_env
    # -------------------------------------------------------------------------
    try:
        actual_headers = [normalize(ws.cell(row=1, column=c).value) for c in range(1, 5)]
        headers_ok = actual_headers == EXPECTED_HEADERS

        data_row_count = ws.max_row - 1  # subtract header row

        if headers_ok and data_row_count == 7:
            print(f"PASS: Component 1 — correct headers {actual_headers} and 7 data rows (0.3 pts)")
            total_score += 0.3
        else:
            if not headers_ok:
                print(f"FAIL: Component 1 — headers mismatch: expected {EXPECTED_HEADERS}, found {actual_headers}")
            if data_row_count != 7:
                print(f"FAIL: Component 1 — expected 7 data rows, found {data_row_count}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: All 7 rows have correct data content (0.5 points)
    # Check Author Name, Email, Benchmark Name, Year for each row
    # Regardless of order — order is checked in Component 3
    # -------------------------------------------------------------------------
    try:
        actual_rows = []
        for row in range(2, ws.max_row + 1):
            author = normalize(ws.cell(row=row, column=1).value)
            email  = normalize(ws.cell(row=row, column=2).value)
            bench  = normalize(ws.cell(row=row, column=3).value)
            year   = year_to_int(ws.cell(row=row, column=4).value)
            actual_rows.append((author, email, bench, year))

        # Build sets for unordered comparison
        expected_set = set()
        for (a, e, b, y) in EXPECTED_DATA:
            expected_set.add((normalize(a), normalize(e), normalize(b), y))

        actual_set = set()
        for (a, e, b, y) in actual_rows:
            actual_set.add((normalize(a), normalize(e), normalize(b), y))

        missing = expected_set - actual_set
        extra   = actual_set - expected_set

        if not missing and not extra:
            print(f"PASS: Component 2 — all 7 rows have correct Author Name, Email, Benchmark Name, Year (0.5 pts)")
            total_score += 0.5
        else:
            if missing:
                print(f"FAIL: Component 2 — missing rows: {missing}")
            if extra:
                print(f"FAIL: Component 2 — extra/incorrect rows: {extra}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Data sorted by Year ascending, then Author Name ascending (0.2 points)
    # Expected order: 2015(Russakovsky), 2016(Rajpurkar), 2018(Wang),
    #                 2019(Zellers), 2021(Hendrycks), 2021(Chen), 2022(Srivastava)
    # -------------------------------------------------------------------------
    try:
        actual_order = []
        for row in range(2, ws.max_row + 1):
            author = normalize(ws.cell(row=row, column=1).value)
            year   = year_to_int(ws.cell(row=row, column=4).value)
            actual_order.append((year, author))

        # Compute expected order from EXPECTED_DATA (already sorted)
        expected_order = [(y, normalize(a)) for (a, e, b, y) in EXPECTED_DATA]

        if actual_order == expected_order:
            print(f"PASS: Component 3 — data correctly sorted by Year then Author Name (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — incorrect sort order")
            print(f"  Expected: {expected_order}")
            print(f"  Actual:   {actual_order}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point: verify the canonical output file on the VM
if not os.path.exists(FILE_PATH):
    print(f"File not found: {FILE_PATH}")
    print("REWARD: 0.0")
else:
    verify_task(FILE_PATH)
