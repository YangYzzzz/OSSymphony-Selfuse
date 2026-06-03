"""
Initial Setup: Team commission rollup workbook with 5 rep sheets and blank CommSummary
Task ID: calc_sales_commission_team_rollup_075
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_commission_team_rollup_075'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def add_rep_sheet(wb, sheet_name, rep_first, rep_last, deals):
    """
    Creates a rep commission sheet with deal data.
    F22 holds the total commission.
    """
    ws = wb.create_sheet(sheet_name)

    # Header row
    headers = ['Deal ID', 'Client', 'Close Date', 'Deal Value', 'Commission Rate', 'Commission']
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 14

    # Title row above data
    ws.insert_rows(1)
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"{rep_first} {rep_last} — Sales Commission Report Q1 2025"
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center")

    # Now data starts at row 3 (headers at row 2)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for r_idx, deal in enumerate(deals, 3):
        deal_id, client, close_date, deal_value, comm_rate = deal
        commission = round(deal_value * comm_rate, 2)

        ws.cell(row=r_idx, column=1, value=deal_id)
        ws.cell(row=r_idx, column=2, value=client)
        ws.cell(row=r_idx, column=3, value=close_date)
        cell_d = ws.cell(row=r_idx, column=4, value=deal_value)
        cell_d.number_format = '$#,##0.00'
        cell_e = ws.cell(row=r_idx, column=5, value=comm_rate)
        cell_e.number_format = '0.00%'
        cell_f = ws.cell(row=r_idx, column=6, value=commission)
        cell_f.number_format = '$#,##0.00'

    # Row 22 (data rows are 3-21, so row 22 is the total row)
    # We need F22 to be the total commission
    # Data rows: 3 to 3+len(deals)-1, which is rows 3..22 for 20 deals
    total_row = 22
    ws.cell(row=total_row, column=5, value="TOTAL").font = Font(bold=True)
    total_cell = ws.cell(row=total_row, column=6, value=f"=SUM(F3:F{total_row - 1})")
    total_cell.number_format = '$#,##0.00'
    total_cell.font = Font(bold=True)

    # Subtotal label in col A
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)

    return ws


def create_initial():
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_ws = wb.active
    wb.remove(default_ws)

    # --- Rep: Alex Chen ---
    chen_deals = [
        ('D-2025-001', 'Meridian Healthcare Solutions', '2025-01-08', 48500.00, 0.08),
        ('D-2025-002', 'Apex Technologies Ltd', '2025-01-15', 72300.00, 0.07),
        ('D-2025-003', 'BlueStar Manufacturing', '2025-01-22', 31200.00, 0.09),
        ('D-2025-004', 'Coastal Capital Partners', '2025-02-03', 55800.00, 0.07),
        ('D-2025-005', 'Vanguard Logistics Inc', '2025-02-10', 29400.00, 0.08),
        ('D-2025-006', 'Pinnacle Retail Group', '2025-02-18', 41000.00, 0.07),
        ('D-2025-007', 'Summit Software Corp', '2025-02-24', 63500.00, 0.08),
        ('D-2025-008', 'Harbor View Properties', '2025-03-03', 37800.00, 0.09),
        ('D-2025-009', 'GreenField Energy', '2025-03-11', 52100.00, 0.07),
        ('D-2025-010', 'NextGen Pharma', '2025-03-17', 44600.00, 0.08),
        ('D-2025-011', 'Ironclad Security Services', '2025-03-20', 28300.00, 0.09),
        ('D-2025-012', 'Sterling Financial Advisors', '2025-03-25', 39700.00, 0.07),
        ('D-2025-013', 'Pacific Rim Trading Co', '2025-03-27', 60400.00, 0.08),
        ('D-2025-014', 'Cedarwood Hospitality', '2025-03-29', 33100.00, 0.09),
        ('D-2025-015', 'Orion Digital Media', '2025-03-30', 47200.00, 0.07),
        ('D-2025-016', 'TerraForm Construction', '2025-03-31', 26800.00, 0.08),
        ('D-2025-017', 'Lakeview Medical Center', '2025-03-31', 35600.00, 0.07),
        ('D-2025-018', 'Crimson Analytics', '2025-03-31', 41900.00, 0.08),
        ('D-2025-019', 'Golden Gate Ventures', '2025-03-31', 58300.00, 0.07),
    ]
    add_rep_sheet(wb, 'Chen_Comm', 'Alex', 'Chen', chen_deals)

    # --- Rep: Maria Torres ---
    torres_deals = [
        ('D-2025-020', 'DataStream Solutions', '2025-01-06', 67800.00, 0.08),
        ('D-2025-021', 'Westbrook Insurance Co', '2025-01-14', 43200.00, 0.07),
        ('D-2025-022', 'Northern Lights Brewing', '2025-01-20', 29700.00, 0.09),
        ('D-2025-023', 'Falcon Aerospace Parts', '2025-02-05', 81400.00, 0.07),
        ('D-2025-024', 'Redwood Environmental', '2025-02-12', 36500.00, 0.08),
        ('D-2025-025', 'Clearwater Tech Systems', '2025-02-19', 52800.00, 0.07),
        ('D-2025-026', 'Magnolia Event Planning', '2025-02-26', 24300.00, 0.09),
        ('D-2025-027', 'Iron Peak Mining Ltd', '2025-03-04', 74100.00, 0.08),
        ('D-2025-028', 'Riviera Luxury Hotels', '2025-03-12', 49500.00, 0.07),
        ('D-2025-029', 'Atomic Power Solutions', '2025-03-18', 31800.00, 0.08),
        ('D-2025-030', 'Keystone Financial Group', '2025-03-22', 58700.00, 0.07),
        ('D-2025-031', 'SunRise Agriculture', '2025-03-26', 27400.00, 0.09),
        ('D-2025-032', 'Blueridge Biotech', '2025-03-28', 63200.00, 0.08),
        ('D-2025-033', 'Metro Transit Authority', '2025-03-29', 45600.00, 0.07),
        ('D-2025-034', 'Stellar Gaming Corp', '2025-03-30', 38900.00, 0.08),
        ('D-2025-035', 'Verdant Landscaping', '2025-03-31', 22100.00, 0.09),
        ('D-2025-036', 'Northgate Publishing', '2025-03-31', 41300.00, 0.07),
        ('D-2025-037', 'Cascade Robotics', '2025-03-31', 77600.00, 0.08),
        ('D-2025-038', 'Elite Staffing Agency', '2025-03-31', 33400.00, 0.07),
    ]
    add_rep_sheet(wb, 'Torres_Comm', 'Maria', 'Torres', torres_deals)

    # --- Rep: Jason Liu ---
    liu_deals = [
        ('D-2025-039', 'Horizon Shipping Ltd', '2025-01-09', 58900.00, 0.08),
        ('D-2025-040', 'Crystal Lake Resorts', '2025-01-17', 34700.00, 0.07),
        ('D-2025-041', 'Titan Auto Parts', '2025-01-23', 46200.00, 0.09),
        ('D-2025-042', 'Maple Grove Pharmacy', '2025-02-06', 28800.00, 0.07),
        ('D-2025-043', 'CloudNine Software', '2025-02-13', 71500.00, 0.08),
        ('D-2025-044', 'Bayshore Investment Bank', '2025-02-20', 93200.00, 0.07),
        ('D-2025-045', 'WildCraft Outdoor Gear', '2025-02-27', 39300.00, 0.09),
        ('D-2025-046', 'Delta Precision Mfg', '2025-03-05', 52600.00, 0.08),
        ('D-2025-047', 'Lotus Health Clinic', '2025-03-13', 31100.00, 0.07),
        ('D-2025-048', 'Global Edge Consulting', '2025-03-19', 64800.00, 0.08),
        ('D-2025-049', 'Timber Wolf Lumber', '2025-03-23', 43500.00, 0.07),
        ('D-2025-050', 'Sapphire Digital Agency', '2025-03-27', 26900.00, 0.09),
        ('D-2025-051', 'Crestview Capital', '2025-03-28', 78400.00, 0.08),
        ('D-2025-052', 'Sunrise Foods Group', '2025-03-29', 36700.00, 0.07),
        ('D-2025-053', 'Quantum Security Systems', '2025-03-30', 55300.00, 0.08),
        ('D-2025-054', 'Bridgegate Engineering', '2025-03-31', 42100.00, 0.07),
        ('D-2025-055', 'Moonstone Wellness', '2025-03-31', 29400.00, 0.09),
        ('D-2025-056', 'Skyline Architecture', '2025-03-31', 67800.00, 0.08),
        ('D-2025-057', 'Emerald Coast Media', '2025-03-31', 38200.00, 0.07),
    ]
    add_rep_sheet(wb, 'Liu_Comm', 'Jason', 'Liu', liu_deals)

    # --- Rep: Sandra Park ---
    park_deals = [
        ('D-2025-058', 'Ironwood Capital Partners', '2025-01-10', 42300.00, 0.08),
        ('D-2025-059', 'BlueSky Airlines', '2025-01-16', 89500.00, 0.07),
        ('D-2025-060', 'Westside Medical Group', '2025-01-24', 35600.00, 0.09),
        ('D-2025-061', 'Frontier Electronics', '2025-02-07', 61800.00, 0.07),
        ('D-2025-062', 'Rosewood Hotels & Resorts', '2025-02-14', 47200.00, 0.08),
        ('D-2025-063', 'Alpine Sports Equipment', '2025-02-21', 28900.00, 0.07),
        ('D-2025-064', 'Nexus Data Centers', '2025-02-28', 76300.00, 0.09),
        ('D-2025-065', 'SilverLeaf Publishing', '2025-03-06', 33700.00, 0.08),
        ('D-2025-066', 'Crown Legal Services', '2025-03-14', 54200.00, 0.07),
        ('D-2025-067', 'Velocity Transportation', '2025-03-20', 69400.00, 0.08),
        ('D-2025-068', 'Harborlight Marine', '2025-03-24', 41800.00, 0.07),
        ('D-2025-069', 'Goldfield Mining Corp', '2025-03-28', 55600.00, 0.09),
        ('D-2025-070', 'Cascade Energy Services', '2025-03-29', 32400.00, 0.08),
        ('D-2025-071', 'Bright Horizons Academy', '2025-03-30', 48100.00, 0.07),
        ('D-2025-072', 'Pinnacle Auto Dealers', '2025-03-30', 37500.00, 0.08),
        ('D-2025-073', 'Onyx Tech Ventures', '2025-03-31', 62700.00, 0.07),
        ('D-2025-074', 'BlueBell Healthcare', '2025-03-31', 44300.00, 0.09),
        ('D-2025-075', 'TerraVerde Farms', '2025-03-31', 29800.00, 0.08),
        ('D-2025-076', 'Meridian IT Consulting', '2025-03-31', 53100.00, 0.07),
    ]
    add_rep_sheet(wb, 'Park_Comm', 'Sandra', 'Park', park_deals)

    # --- Rep: Derek Green ---
    green_deals = [
        ('D-2025-077', 'Coastal Ridge Realty', '2025-01-11', 74200.00, 0.08),
        ('D-2025-078', 'Starpoint Financial', '2025-01-18', 38900.00, 0.07),
        ('D-2025-079', 'Phoenix Design Studio', '2025-01-25', 27600.00, 0.09),
        ('D-2025-080', 'Lakeside Breweries', '2025-02-08', 56400.00, 0.07),
        ('D-2025-081', 'Arcadia Systems Inc', '2025-02-15', 83100.00, 0.08),
        ('D-2025-082', 'MetroWide Insurance', '2025-02-22', 45700.00, 0.07),
        ('D-2025-083', 'Redstone Power Corp', '2025-03-01', 32800.00, 0.09),
        ('D-2025-084', 'Neptune Marine Products', '2025-03-07', 61500.00, 0.08),
        ('D-2025-085', 'Emerald Forest Timber', '2025-03-15', 43900.00, 0.07),
        ('D-2025-086', 'Summit View Properties', '2025-03-21', 79200.00, 0.08),
        ('D-2025-087', 'Glacier Point Logistics', '2025-03-25', 37300.00, 0.07),
        ('D-2025-088', 'Moonrise Entertainment', '2025-03-28', 52800.00, 0.09),
        ('D-2025-089', 'Sunridge Solar Energy', '2025-03-29', 68400.00, 0.08),
        ('D-2025-090', 'Harbor Bay Trading', '2025-03-30', 31100.00, 0.07),
        ('D-2025-091', 'Cypress IT Solutions', '2025-03-30', 47600.00, 0.08),
        ('D-2025-092', 'Avalon Aerospace', '2025-03-31', 95300.00, 0.07),
        ('D-2025-093', 'Willow Springs Winery', '2025-03-31', 24700.00, 0.09),
        ('D-2025-094', 'Irongate Construction', '2025-03-31', 58900.00, 0.08),
        ('D-2025-095', 'Pacific Vista Hotels', '2025-03-31', 42100.00, 0.07),
    ]
    add_rep_sheet(wb, 'Green_Comm', 'Derek', 'Green', green_deals)

    # --- CommSummary: blank placeholder sheet ---
    ws_summary = wb.create_sheet('CommSummary')
    # Intentionally left blank — agent must populate this

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: Chen_Comm, Torres_Comm, Liu_Comm, Park_Comm, Green_Comm, CommSummary')
    print('Each rep sheet has total commission in F22')
    print('CommSummary is blank — agent task is to build the rollup')


create_initial()
