"""
Reward Script: Build a team commission rollup report with cross-sheet references
Task ID: calc_sales_commission_team_rollup_075
Domain: libreoffice_calc
Scoring:
  Component 1: CommSummary header, title row, and column headers present (0.20 pts)
  Component 2: Cross-sheet references in B4:B8 pointing to rep commission sheets F22 (0.30 pts)
  Component 3: ROUND formulas in column C for all data rows (0.20 pts)
  Component 4: Team total, manager bonus (5%), and total payroll formulas in rows 9-11 (0.20 pts)
  Component 5: Currency number format ($#,##0.00) and green tab color on CommSummary (0.10 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_commission_team_rollup_075'

REP_SHEETS = ['Chen_Comm', 'Torres_Comm', 'Liu_Comm', 'Park_Comm', 'Green_Comm']
# Expected cross-sheet reference patterns for B4:B8
EXPECTED_REFS = [
    ("='Chen_Comm'!F22", "Chen_Comm"),
    ("='Torres_Comm'!F22", "Torres_Comm"),
    ("='Liu_Comm'!F22", "Liu_Comm"),
    ("='Park_Comm'!F22", "Park_Comm"),
    ("='Green_Comm'!F22", "Green_Comm"),
]


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Gate: CommSummary sheet must exist
    if 'CommSummary' not in wb.sheetnames:
        print("FAIL: 'CommSummary' sheet not found in workbook")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws = wb['CommSummary']

    # Component 1: Title row, column headers, and rep names present (0.20 points)
    # The initial file has CommSummary as a blank sheet — all of these cells must be populated
    try:
        comp1_score = 0.0

        # Check A1 title
        a1_val = ws['A1'].value
        if a1_val and 'Commission Summary' in str(a1_val):
            comp1_score += 0.05
            print(f"PASS: A1 title present: {repr(a1_val)}")
        else:
            print(f"FAIL: A1 should contain 'Commission Summary', found: {repr(a1_val)}")

        # Check column headers in row 3
        headers_found = 0
        expected_headers = ['Rep Name', 'Total Commission', 'Rounded Commission']
        for col_idx, expected_header in enumerate(expected_headers, 1):
            cell_val = ws.cell(row=3, column=col_idx).value
            if cell_val and expected_header.lower() in str(cell_val).lower():
                headers_found += 1
            else:
                print(f"FAIL: Row 3 col {col_idx} expected '{expected_header}', found: {repr(cell_val)}")
        if headers_found == 3:
            comp1_score += 0.05
            print(f"PASS: All 3 column headers found in row 3")
        elif headers_found >= 2:
            comp1_score += 0.03
            print(f"PARTIAL: {headers_found}/3 column headers found in row 3")

        # Check 5 rep names in A4:A8
        rep_names_found = 0
        for row_idx in range(4, 9):
            cell_val = ws.cell(row=row_idx, column=1).value
            if cell_val and str(cell_val).strip():
                rep_names_found += 1
        if rep_names_found == 5:
            comp1_score += 0.10
            print(f"PASS: All 5 rep names found in A4:A8")
        elif rep_names_found >= 3:
            comp1_score += 0.05
            print(f"PARTIAL: {rep_names_found}/5 rep names found in A4:A8")
        else:
            print(f"FAIL: Only {rep_names_found}/5 rep names found in A4:A8")

        total_score += comp1_score
        print(f"Component 1 score: {comp1_score}/0.20")

    except Exception as e:
        print(f"ERROR: Component 1 check failed: {e}")

    # Component 2: Cross-sheet references in B4:B8 pointing to rep commission sheets F22 (0.30 points)
    # Each cross-sheet reference earns 0.06 pts (5 reps * 0.06 = 0.30)
    try:
        comp2_score = 0.0
        refs_correct = 0

        for row_idx, (expected_ref, sheet_name) in enumerate(EXPECTED_REFS, 4):
            cell_val = ws.cell(row=row_idx, column=2).value
            if cell_val and isinstance(cell_val, str):
                # Normalize: remove extra spaces, check case-insensitive
                normalized = cell_val.strip().replace(' ', '')
                expected_normalized = expected_ref.replace(' ', '')
                # Also accept variations like ='Chen_Comm'!F22 or =Chen_Comm!F22
                # Check that it references the right sheet and cell F22
                has_sheet_ref = sheet_name.lower() in cell_val.lower()
                has_f22 = 'F22' in cell_val.upper() or 'f22' in cell_val.lower()
                is_formula = cell_val.startswith('=')
                if is_formula and has_sheet_ref and has_f22:
                    refs_correct += 1
                    print(f"PASS: B{row_idx} has cross-sheet ref to {sheet_name}!F22: {repr(cell_val)}")
                else:
                    print(f"FAIL: B{row_idx} expected cross-sheet ref to {sheet_name}!F22, found: {repr(cell_val)}")
            else:
                print(f"FAIL: B{row_idx} expected formula, found: {repr(cell_val)}")

        comp2_score = refs_correct * 0.06
        total_score += comp2_score
        print(f"Component 2 score: {comp2_score}/0.30 ({refs_correct}/5 cross-sheet refs correct)")

    except Exception as e:
        print(f"ERROR: Component 2 check failed: {e}")

    # Component 3: ROUND formulas in column C for rep rows (C4:C8), team total (C9), etc. (0.20 points)
    # C4:C8 and C9 should use ROUND formulas
    try:
        comp3_score = 0.0
        round_correct = 0
        # Check C4:C9 (6 cells, 0.20/6 each ~ 0.033 each)
        round_cells = [(4, 'B4'), (5, 'B5'), (6, 'B6'), (7, 'B7'), (8, 'B8'), (9, 'B9')]
        for row_idx, b_ref in round_cells:
            cell_val = ws.cell(row=row_idx, column=3).value
            if cell_val and isinstance(cell_val, str):
                # Should be =ROUND(Bx, 2) or =ROUND(Bx,2)
                normalized = cell_val.strip().upper().replace(' ', '')
                expected_pattern = f'=ROUND({b_ref.upper()},2)'
                if normalized == expected_pattern:
                    round_correct += 1
                    print(f"PASS: C{row_idx} has correct ROUND formula: {repr(cell_val)}")
                elif 'ROUND' in normalized and b_ref.upper() in normalized and '2' in normalized:
                    round_correct += 1
                    print(f"PASS: C{row_idx} has ROUND formula (flexible match): {repr(cell_val)}")
                else:
                    print(f"FAIL: C{row_idx} expected ROUND formula referencing {b_ref}, found: {repr(cell_val)}")
            else:
                print(f"FAIL: C{row_idx} expected ROUND formula, found: {repr(cell_val)}")

        comp3_score = round(round_correct * (0.20 / 6), 4)
        total_score += comp3_score
        print(f"Component 3 score: {comp3_score}/0.20 ({round_correct}/6 ROUND formulas correct)")

    except Exception as e:
        print(f"ERROR: Component 3 check failed: {e}")

    # Component 4: Team total (row 9), manager bonus 5% (row 10), total payroll (row 11) (0.20 points)
    try:
        comp4_score = 0.0

        # Row 9: TEAM TOTAL with SUM formula
        a9_val = ws['A9'].value
        b9_val = ws['B9'].value
        if a9_val and 'TEAM TOTAL' in str(a9_val).upper():
            comp4_score += 0.03
            print(f"PASS: A9 has 'TEAM TOTAL' label: {repr(a9_val)}")
        else:
            print(f"FAIL: A9 expected 'TEAM TOTAL', found: {repr(a9_val)}")

        if b9_val and isinstance(b9_val, str) and b9_val.startswith('='):
            # Should sum B4:B8
            normalized_b9 = b9_val.strip().upper().replace(' ', '')
            if 'SUM' in normalized_b9 and 'B4' in normalized_b9 and 'B8' in normalized_b9:
                comp4_score += 0.07
                print(f"PASS: B9 has SUM formula for team total: {repr(b9_val)}")
            else:
                print(f"FAIL: B9 should be =SUM(B4:B8), found: {repr(b9_val)}")
        else:
            print(f"FAIL: B9 expected SUM formula, found: {repr(b9_val)}")

        # Row 10: MANAGER BONUS = 5% of team total
        a10_val = ws['A10'].value
        b10_val = ws['B10'].value
        if a10_val and 'MANAGER' in str(a10_val).upper() and 'BONUS' in str(a10_val).upper():
            comp4_score += 0.02
            print(f"PASS: A10 has 'MANAGER BONUS' label: {repr(a10_val)}")
        else:
            print(f"FAIL: A10 expected 'MANAGER BONUS', found: {repr(a10_val)}")

        if b10_val and isinstance(b10_val, str) and b10_val.startswith('='):
            normalized_b10 = b10_val.strip().upper().replace(' ', '')
            # Should reference B9 and multiply by 0.05
            if 'B9' in normalized_b10 and ('0.05' in normalized_b10 or '5%' in normalized_b10 or '0.05' in b10_val):
                comp4_score += 0.04
                print(f"PASS: B10 has manager bonus formula (5% of B9): {repr(b10_val)}")
            else:
                print(f"FAIL: B10 should be =B9*0.05 or similar, found: {repr(b10_val)}")
        else:
            print(f"FAIL: B10 expected formula for manager bonus, found: {repr(b10_val)}")

        # Row 11: TOTAL PAYROLL = team total + manager bonus
        a11_val = ws['A11'].value
        b11_val = ws['B11'].value
        if a11_val and 'TOTAL PAYROLL' in str(a11_val).upper():
            comp4_score += 0.02
            print(f"PASS: A11 has 'TOTAL PAYROLL' label: {repr(a11_val)}")
        else:
            print(f"FAIL: A11 expected 'TOTAL PAYROLL', found: {repr(a11_val)}")

        if b11_val and isinstance(b11_val, str) and b11_val.startswith('='):
            normalized_b11 = b11_val.strip().upper().replace(' ', '')
            # Should sum B9 and B10
            if 'B9' in normalized_b11 and 'B10' in normalized_b11:
                comp4_score += 0.02
                print(f"PASS: B11 has total payroll formula (B9+B10): {repr(b11_val)}")
            else:
                print(f"FAIL: B11 should reference B9 and B10, found: {repr(b11_val)}")
        else:
            print(f"FAIL: B11 expected formula for total payroll, found: {repr(b11_val)}")

        total_score += comp4_score
        print(f"Component 4 score: {comp4_score}/0.20")

    except Exception as e:
        print(f"ERROR: Component 4 check failed: {e}")

    # Component 5: Currency format ($#,##0.00) in B and C columns, and green tab color (0.10 points)
    try:
        comp5_score = 0.0

        # Check currency format in B column (B4:B11)
        currency_format_count = 0
        for row_idx in range(4, 12):
            cell = ws.cell(row=row_idx, column=2)
            if cell.number_format and '$' in cell.number_format and '0.00' in cell.number_format:
                currency_format_count += 1
        if currency_format_count >= 7:
            comp5_score += 0.05
            print(f"PASS: Currency format $#,##0.00 found in {currency_format_count}/8 B column cells")
        elif currency_format_count >= 4:
            comp5_score += 0.03
            print(f"PARTIAL: Currency format found in {currency_format_count}/8 B column cells")
        else:
            print(f"FAIL: Currency format found in only {currency_format_count}/8 B column cells")

        # Check green tab color
        try:
            tab_color = ws.sheet_properties.tabColor
            if tab_color is not None:
                rgb = str(tab_color.rgb).upper() if hasattr(tab_color, 'rgb') else ''
                # Green colors: FF00B050 (Excel green), FF00FF00 (pure green), or any green-ish
                is_green = False
                if rgb:
                    # Parse ARGB
                    r_val = int(rgb[2:4], 16) if len(rgb) >= 8 else 0
                    g_val = int(rgb[4:6], 16) if len(rgb) >= 8 else 0
                    b_val = int(rgb[6:8], 16) if len(rgb) >= 8 else 0
                    # Green if green channel is dominant
                    is_green = g_val > r_val and g_val > b_val
                if is_green:
                    comp5_score += 0.05
                    print(f"PASS: CommSummary tab color is green: {rgb}")
                else:
                    print(f"FAIL: CommSummary tab color is not green: {rgb}")
            else:
                print(f"FAIL: CommSummary tab has no color set")
        except Exception as e:
            print(f"ERROR: Tab color check failed: {e}")

        total_score += comp5_score
        print(f"Component 5 score: {comp5_score}/0.10")

    except Exception as e:
        print(f"ERROR: Component 5 check failed: {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
