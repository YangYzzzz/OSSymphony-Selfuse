"""
Initial Setup: Fix text-extraction formula that errors on rows without hyphens
Task ID: calc_tbl_076
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_076'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # --- Headers ---
    headers = ["Product Code", "Extracted Prefix", "Category", "Price"]
    header_font = Font(bold=True, size=11, name="Calibri")
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    white_font = Font(bold=True, size=11, name="Calibri", color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Data rows ---
    # Mix of codes WITH hyphens and WITHOUT hyphens
    data = [
        ["WDG-4417", None, "Widgets", 29.99],
        ["BLT-8823", None, "Bolts", 4.50],
        ["NUT5590", None, "Nuts", 2.75],
        ["SCR-1102", None, "Screws", 3.25],
        ["WSH7764", None, "Washers", 1.80],
        ["FLG-3398", None, "Flanges", 18.50],
        ["BRG2241", None, "Bearings", 42.00],
        ["SPR-6650", None, "Springs", 7.95],
        ["PIN-0087", None, "Pins", 1.20],
        ["RVT3345", None, "Rivets", 0.95],
        ["GKT-5579", None, "Gaskets", 12.30],
        ["CLR8812", None, "Collars", 8.75],
        ["BUS-2204", None, "Bushings", 15.60],
        ["ROD-9931", None, "Rods", 22.40],
        ["CAP4467", None, "Caps", 3.10],
        ["HNG-1158", None, "Hinges", 9.85],
        ["AXL7703", None, "Axles", 35.20],
        ["SHF-3346", None, "Shafts", 28.90],
        ["KEY0092", None, "Keys", 5.45],
        ["GER-8815", None, "Gears", 55.00],
        ["VLV2238", None, "Valves", 67.50],
        ["PMP-4471", None, "Pumps", 125.00],
        ["MTR6609", None, "Motors", 245.00],
        ["SEN-1134", None, "Sensors", 89.99],
        ["RLY5567", None, "Relays", 14.75],
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])  # Product Code
        # Column B: BROKEN formula - uses FIND("-") which errors when no hyphen
        ws.cell(row=r, column=2, value=f'=LEFT(A{r},FIND("-",A{r})-1)')
        ws.cell(row=r, column=3, value=row_data[2])  # Category
        ws.cell(row=r, column=4, value=row_data[3])  # Price

    # Format price column
    for r in range(2, len(data) + 2):
        ws.cell(row=r, column=4).number_format = '$#,##0.00'

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12

    # Freeze header row
    ws.freeze_panes = "A2"

    # --- Instructions sheet ---
    ws2 = wb.create_sheet("Notes")
    ws2["A1"] = "Product Code Format"
    ws2["A1"].font = Font(bold=True, size=13)
    ws2["A3"] = "Some product codes use the format PREFIX-NUMBER (e.g., WDG-4417)."
    ws2["A4"] = "Others omit the hyphen (e.g., NUT5590)."
    ws2["A5"] = "The formula in column B of the Data sheet should extract the prefix."
    ws2["A6"] = "Currently it errors on codes without a hyphen."
    ws2.column_dimensions["A"].width = 65

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
