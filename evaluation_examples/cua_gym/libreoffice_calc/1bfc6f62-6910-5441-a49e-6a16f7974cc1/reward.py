"""
Reward Script: Fix IFERROR formula for text extraction before hyphen
Task ID: calc_tbl_076
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): All B-column formulas wrapped with IFERROR
  Component 2 (0.3): LEFT+FIND extraction logic preserved inside IFERROR
  Component 3 (0.3): IFERROR fallback references full cell value (Ax)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_076'


def persist_app_state(domain):
    """Best-effort save of any open LibreOffice documents."""
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print(f"PERSIST: ctrl+s sent for {domain}")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Data']

    # Determine the data rows (row 2 onwards with data in column A)
    data_rows = []
    for r in range(2, ws.max_row + 1):
        a_val = ws.cell(row=r, column=1).value
        if a_val is not None:
            data_rows.append(r)

    if not data_rows:
        print("FAIL: No data rows found in column A")
        print("REWARD: 0.0")
        return 0.0

    print(f"INFO: Found {len(data_rows)} data rows: {data_rows[0]}-{data_rows[-1]}")

    # Component 1: All B-column formulas use IFERROR wrapping (0.4 points)
    # This is the CORE change: initial uses bare LEFT(Ax,FIND(...)), golden wraps with IFERROR
    try:
        iferror_count = 0
        total_formula_count = 0
        for r in data_rows:
            b_val = ws.cell(row=r, column=2).value
            if b_val is not None and isinstance(b_val, str) and b_val.startswith('='):
                total_formula_count += 1
                normalized = b_val.upper().replace(" ", "")
                if normalized.startswith("=IFERROR("):
                    iferror_count += 1

        if total_formula_count == 0:
            print("FAIL: Component 1 -- No formulas found in column B")
        elif iferror_count == total_formula_count:
            print(f"PASS: Component 1 -- All {iferror_count}/{total_formula_count} B-column formulas use IFERROR (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 -- Only {iferror_count}/{total_formula_count} formulas use IFERROR")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: IFERROR formulas preserve LEFT+FIND extraction logic inside (0.3 points)
    # Compound check: must have IFERROR AND contain LEFT(Ax,FIND("-",Ax)-1) extraction
    # This anchors to the task change -- bare LEFT+FIND without IFERROR does NOT score
    try:
        compound_pass_count = 0
        checked_count = 0
        for r in data_rows:
            b_val = ws.cell(row=r, column=2).value
            if b_val is not None and isinstance(b_val, str) and b_val.startswith('='):
                checked_count += 1
                normalized = b_val.upper().replace(" ", "")
                has_iferror = normalized.startswith("=IFERROR(")
                has_left_find = "LEFT(" in normalized and "FIND(" in normalized
                if has_iferror and has_left_find:
                    compound_pass_count += 1

        if checked_count == 0:
            print("FAIL: Component 2 -- No formulas to check")
        elif compound_pass_count == checked_count:
            print(f"PASS: Component 2 -- All {compound_pass_count}/{checked_count} formulas have IFERROR + LEFT+FIND (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 -- Only {compound_pass_count}/{checked_count} formulas have IFERROR wrapping LEFT+FIND")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: IFERROR fallback references the cell in column A (0.3 points)
    # e.g., =IFERROR(LEFT(A2,FIND("-",A2)-1),A2) -- the second arg of IFERROR should be Ax
    try:
        fallback_correct_count = 0
        checked_count = 0
        for r in data_rows:
            b_val = ws.cell(row=r, column=2).value
            if b_val is not None and isinstance(b_val, str) and b_val.startswith('='):
                checked_count += 1
                normalized = b_val.upper().replace(" ", "")
                # Extract the IFERROR second argument: should be A<row>
                # Pattern: =IFERROR(...,A<row>)
                expected_ref = f"A{r}"
                # Check if formula ends with ,A<row>)
                if normalized.startswith("=IFERROR(") and normalized.endswith(f",{expected_ref})"):
                    fallback_correct_count += 1

        if checked_count == 0:
            print("FAIL: Component 3 -- No formulas to check")
        elif fallback_correct_count == checked_count:
            print(f"PASS: Component 3 -- All {fallback_correct_count}/{checked_count} formulas have correct Ax fallback (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 3 -- Only {fallback_correct_count}/{checked_count} formulas have correct Ax fallback reference")
            # Show a sample failure for debugging
            for r in data_rows:
                b_val = ws.cell(row=r, column=2).value
                if b_val is not None and isinstance(b_val, str) and b_val.startswith('='):
                    normalized = b_val.upper().replace(" ", "")
                    expected_ref = f"A{r}"
                    if not (normalized.startswith("=IFERROR(") and normalized.endswith(f",{expected_ref})")):
                        print(f"  Sample: B{r} = {b_val!r} (expected fallback to {expected_ref})")
                        break
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
