"""
Initial Setup: Create unformatted inventory spreadsheet
Task ID: calc_gg2_038
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg2_038'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventory'

    # Headers (A1:H1)
    headers = ['SKU', 'Name', 'Category', 'Supplier', 'Unit Cost',
               'Qty On Hand', 'Reorder Point', 'Status']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 24 rows of realistic inventory data (rows 2-25)
    data = [
        ['WH-1001', 'Industrial Ball Bearing 6205', 'Bearings', 'Pacific Components Ltd', 12.45, 350, 100, 'In Stock'],
        ['WH-1002', 'Stainless Steel Hex Bolt M10x40', 'Fasteners', 'Summit Hardware Co', 0.85, 2400, 500, 'In Stock'],
        ['WH-1003', 'Hydraulic Cylinder Seal Kit', 'Seals', 'FluidTech Industries', 34.90, 45, 20, 'In Stock'],
        ['WH-1004', 'Carbon Steel Flat Washer 3/8"', 'Fasteners', 'Summit Hardware Co', 0.12, 8500, 2000, 'In Stock'],
        ['WH-1005', 'Pneumatic Air Filter Element', 'Filters', 'AirFlow Systems Inc', 28.75, 12, 25, 'Low Stock'],
        ['WH-1006', 'Nylon Cable Tie 300mm Black', 'Electrical', 'Vertex Supply Group', 0.08, 15000, 3000, 'In Stock'],
        ['WH-1007', 'Copper Pipe Fitting 1/2" Elbow', 'Plumbing', 'MetalWorks Direct', 3.20, 620, 150, 'In Stock'],
        ['WH-1008', 'High-Temp Silicone Gasket RTV', 'Adhesives', 'ChemBond Solutions', 9.50, 88, 30, 'In Stock'],
        ['WH-1009', 'Linear Motion Guide Rail 400mm', 'Motion', 'Pacific Components Ltd', 145.00, 8, 5, 'In Stock'],
        ['WH-1010', 'Polyurethane V-Belt A68', 'Belts', 'DriveLink Corp', 18.30, 0, 10, 'Out of Stock'],
        ['WH-1011', 'Aluminum Extrusion T-Slot 2040', 'Structural', 'MetalWorks Direct', 22.60, 175, 50, 'In Stock'],
        ['WH-1012', 'Miniature Circuit Breaker 16A', 'Electrical', 'Vertex Supply Group', 7.40, 310, 100, 'In Stock'],
        ['WH-1013', 'PTFE Thread Seal Tape 12mm', 'Plumbing', 'ChemBond Solutions', 1.95, 540, 200, 'In Stock'],
        ['WH-1014', 'Deep Groove Ball Bearing 6308', 'Bearings', 'Pacific Components Ltd', 24.80, 92, 40, 'In Stock'],
        ['WH-1015', 'Stainless Steel Socket Cap M8x25', 'Fasteners', 'Summit Hardware Co', 0.65, 3200, 800, 'In Stock'],
        ['WH-1016', 'Activated Carbon Filter Cartridge', 'Filters', 'AirFlow Systems Inc', 42.00, 5, 15, 'Low Stock'],
        ['WH-1017', 'Heat Shrink Tubing Kit Assorted', 'Electrical', 'Vertex Supply Group', 15.90, 67, 25, 'In Stock'],
        ['WH-1018', 'Brass Ball Valve 3/4"', 'Plumbing', 'MetalWorks Direct', 19.75, 148, 50, 'In Stock'],
        ['WH-1019', 'Timing Belt HTD 5M 300mm', 'Belts', 'DriveLink Corp', 31.40, 22, 10, 'In Stock'],
        ['WH-1020', 'Spring Lock Washer M12', 'Fasteners', 'Summit Hardware Co', 0.18, 6700, 1500, 'In Stock'],
        ['WH-1021', 'Precision Linear Shaft 16mm', 'Motion', 'Pacific Components Ltd', 56.00, 14, 8, 'In Stock'],
        ['WH-1022', 'Industrial Epoxy Adhesive 50ml', 'Adhesives', 'ChemBond Solutions', 13.25, 3, 10, 'Low Stock'],
        ['WH-1023', 'DIN Rail 35mm Standard 1m', 'Structural', 'MetalWorks Direct', 4.80, 230, 60, 'In Stock'],
        ['WH-1024', 'Flexible Conduit 20mm PVC Grey', 'Electrical', 'Vertex Supply Group', 2.35, 890, 200, 'In Stock'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths for readability (no styling/formatting)
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
