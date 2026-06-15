"""
Reward Script: Apply AutoFormat to inventory table then override header row font
Task ID: calc_gg2_038
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Header row A1:H1 is bold
  Component 2 (0.25): Header row A1:H1 is Calibri 11pt
  Component 3 (0.25): Data rows have AutoFormat font (different from default Calibri/11)
  Component 4 (0.25): Alternating/banded row fill colors on data rows (AutoFormat styling)
"""

import os

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg2_038'


def persist_app_state(domain: str):
    """Try to save any unsaved state in LibreOffice."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify 'Inventory' sheet exists
    if 'Inventory' not in wb.sheetnames:
        print("CRITICAL: 'Inventory' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Inventory']

    # Component 1: Header row A1:H1 is bold (0.25 points)
    # Initial state: bold=False. Golden state: bold=True.
    try:
        bold_count = 0
        for col in range(1, 9):
            cell = ws.cell(row=1, column=col)
            if cell.font.bold:
                bold_count += 1
        if bold_count == 8:
            print(f"PASS: Component 1 — All 8 header cells are bold (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — {bold_count}/8 header cells are bold (need all 8)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Header row A1:H1 has Calibri font at 11pt (0.25 points)
    # This verifies the font override was applied (Calibri 11pt specifically).
    # In the initial state, font is also Calibri 11pt BUT not bold — so this component
    # alone would pass on initial. It is gated by Component 1 (bold check) for discrimination.
    # However, to be safe, we combine: font must be Calibri AND 11pt AND bold (the override).
    try:
        calibri_11_bold_count = 0
        for col in range(1, 9):
            cell = ws.cell(row=1, column=col)
            font_name = cell.font.name
            font_size = cell.font.size
            is_bold = cell.font.bold
            if font_name == 'Calibri' and font_size == 11.0 and is_bold:
                calibri_11_bold_count += 1
            else:
                print(f"  INFO: {cell.coordinate} font={font_name}/{font_size}/bold={is_bold}")
        if calibri_11_bold_count == 8:
            print(f"PASS: Component 2 — All 8 header cells are Calibri 11pt Bold (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — {calibri_11_bold_count}/8 header cells match Calibri 11pt Bold")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Data rows have AutoFormat font applied (0.25 points)
    # Initial state: all data rows have Calibri/11.0 (default). After AutoFormat, data rows
    # should have a DIFFERENT font (e.g., Liberation Sans/10.0) since AutoFormat changes the font.
    # The header font was overridden back to Calibri, but data rows retain AutoFormat font.
    # We check that data rows do NOT have the initial default (Calibri/11.0).
    try:
        changed_font_count = 0
        total_data_cells = 0
        for row in range(2, 26):
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                total_data_cells += 1
                font_name = cell.font.name
                font_size = cell.font.size
                # If font differs from default Calibri/11, AutoFormat was applied
                if font_name != 'Calibri' or font_size != 11.0:
                    changed_font_count += 1

        # Require at least 80% of data cells to have non-default font
        threshold = int(total_data_cells * 0.8)
        if changed_font_count >= threshold:
            print(f"PASS: Component 3 — {changed_font_count}/{total_data_cells} data cells have AutoFormat font (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 3 — Only {changed_font_count}/{total_data_cells} data cells have non-default font (need >= {threshold})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Alternating/banded row fill colors on data rows (0.25 points)
    # Initial state: no fill on any data row. After AutoFormat, even-numbered data rows
    # should have a solid fill color (banding pattern).
    # We check that at least some rows have solid fills and some don't (alternating).
    try:
        filled_rows = 0
        unfilled_rows = 0
        for row in range(2, 26):
            cell = ws.cell(row=row, column=1)  # Check column A as representative
            fill_type = cell.fill.patternType
            fill_color = cell.fill.fgColor.rgb if cell.fill.fgColor else None
            has_fill = (fill_type == 'solid' and fill_color not in (None, '00000000'))
            if has_fill:
                filled_rows += 1
            else:
                unfilled_rows += 1

        # For banding, we need both filled and unfilled rows
        # With 24 data rows, alternating means ~12 each
        if filled_rows >= 6 and unfilled_rows >= 6:
            print(f"PASS: Component 4 — Alternating fill: {filled_rows} filled, {unfilled_rows} unfilled data rows (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 4 — Not enough alternation: {filled_rows} filled, {unfilled_rows} unfilled (need >= 6 each)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
