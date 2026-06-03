"""
Initial Setup: Create a correlation matrix spreadsheet for visualization task
Task ID: calc_gcp_071
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_071'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Correlations'

    # Variable names
    variables = [
        'Sales',
        'Marketing Spend',
        'Headcount',
        'Customer Satisfaction',
        'Website Traffic',
        'Social Followers',
    ]

    # 6x6 correlation matrix (symmetric, diagonal=1.00)
    # Realistic correlation values
    matrix = [
        [1.00,  0.85,  0.62,  0.74,  0.91, 0.53],
        [0.85,  1.00,  0.48,  0.67,  0.78, 0.41],
        [0.62,  0.48,  1.00,  0.29,  0.55, -0.32],
        [0.74,  0.67,  0.29,  1.00,  0.68, 0.44],
        [0.91,  0.78,  0.55,  0.68,  1.00, 0.50],
        [0.53,  0.41, -0.32,  0.44,  0.50, 1.00],
    ]

    # A1: corner label
    ws.cell(row=1, column=1, value='Variables')
    ws['A1'].font = Font(bold=True, size=11)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Header row (B1:G1) - variable names
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, var_name in enumerate(variables, 2):
        cell = ws.cell(row=1, column=col_idx, value=var_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Row labels (A2:A7) - variable names
    label_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    label_font = Font(bold=True, color="FFFFFF", size=11)
    for row_idx, var_name in enumerate(variables, 2):
        cell = ws.cell(row=row_idx, column=1, value=var_name)
        cell.font = label_font
        cell.fill = label_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')

    # Fill matrix values (B2:G7)
    for r_idx, row_data in enumerate(matrix):
        for c_idx, val in enumerate(row_data):
            cell = ws.cell(row=r_idx + 2, column=c_idx + 2, value=val)
            cell.number_format = '0.00'
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Color code: diagonal=gold, positive=green shades, negative=red shades
            if r_idx == c_idx:
                cell.fill = PatternFill(start_color="FFFFD700", end_color="FFFFD700", fill_type="solid")
                cell.font = Font(bold=True, size=11)
            elif val < 0:
                cell.fill = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid")
            elif val >= 0.7:
                cell.fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")

    # Column widths
    ws.column_dimensions['A'].width = 22
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].width = 18

    # Row heights
    ws.row_dimensions[1].height = 35
    for r in range(2, 8):
        ws.row_dimensions[r].height = 25

    # Add thin borders to the matrix area
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(1, 8):
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = border

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
