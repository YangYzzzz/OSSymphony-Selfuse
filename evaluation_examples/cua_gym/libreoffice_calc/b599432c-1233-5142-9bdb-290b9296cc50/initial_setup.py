"""
Initial Setup: HR Succession Planning Matrix
Task ID: calc_hr_succession_planning_068
Domain: libreoffice_calc

Creates a succession planning spreadsheet with key roles and candidates.
Columns D, F, G have inconsistent or empty values (NO data validation dropdowns).
NO conditional formatting applied.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_succession_planning_068'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Succession Plan'

    # --- Row 1: Headers ---
    headers = ['Role', 'Current Incumbent', 'Successor 1', 'Readiness',
               'Successor 2', 'Readiness (S2)', 'Risk Level']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Rows 2-28: Key roles and succession candidates ---
    # Columns: Role, Current Incumbent, Successor 1, Readiness, Successor 2, Readiness (S2), Risk Level
    # NOTE: D (Readiness), F (Readiness S2), G (Risk Level) are intentionally inconsistent/incomplete
    # to represent real-world data that needs cleanup — no dropdowns, no conditional formatting
    data = [
        ('Chief Executive Officer',       'Margaret Sullivan',   'David Okafor',       'Ready Now',    'Priya Nair',          'N/A',       'Critical'),
        ('Chief Financial Officer',        'Robert Hensley',      'Aisha Kamara',       '1-2 Years',    'Thomas Brandt',       '3+ Years',  'Critical'),
        ('Chief Operating Officer',        'Linda Yamamoto',      'Carlos Estrada',     '',             '',                    '',          ''),
        ('Chief Technology Officer',       'James Whitfield',     'Mei-Ling Zhao',      '3+ Years',     '',                    '',          'High'),
        ('Chief Human Resources Officer',  'Sandra Kowalski',     'Omar Farouk',        'Ready Now',    'Natasha Ivanova',     '1-2 Years', 'High'),
        ('VP Sales & Marketing',           'Brian Nguyen',        'Fatima Al-Hassan',   '1-2 Years',    'Patrick O\'Brien',    '3+ Years',  'Critical'),
        ('VP Engineering',                 'Alicia Fernandez',    'Ravi Subramaniam',   '',             'Yuki Tanaka',         '',          'Critical'),
        ('VP Operations',                  'Derek Chambers',      '',                   'No Successor', '',                    'N/A',       'Critical'),
        ('VP Finance',                     'Heather MacAllister', 'Leon Dubois',        '3+ Years',     '',                    '',          'High'),
        ('VP Legal & Compliance',          'Gregory Holloway',    '',                   'No Successor', '',                    'N/A',       'Critical'),
        ('Director of Product Management', 'Emily Chen',          'Marcus Williams',    '1-2 Years',    'Sophie Bergmann',     '1-2 Years', 'High'),
        ('Director of Data Science',       'Kwame Asante',        'Isabella Romano',    '3+ Years',     '',                    '',          'High'),
        ('Director of Cybersecurity',      'Nathan Patel',        '',                   'No Successor', '',                    'N/A',       'Critical'),
        ('Director of Supply Chain',       'Victoria Johansson',  'Hiroshi Nakamura',   '1-2 Years',    'Amara Diallo',        '3+ Years',  'Medium'),
        ('Director of Customer Success',   'Antonio Reyes',       'Clara Hoffmann',     'Ready Now',    'Adeola Adeyemi',      'Ready Now', 'Medium'),
        ('Director of Finance',            'Bridget Murphy',      'Samuel Lee',         '1-2 Years',    '',                    '',          'High'),
        ('Director of HR Business Partner','Chidi Okonkwo',       'Laura Santos',       '3+ Years',     '',                    '',          'Medium'),
        ('Head of Strategic Planning',     'Samantha Blake',      'Eric Lindqvist',     '1-2 Years',    'Nadia Petrov',        '1-2 Years', 'High'),
        ('Head of Internal Audit',         'Franklin Torres',     '',                   'No Successor', '',                    'N/A',       'Critical'),
        ('Head of Corporate Dev',          'Yolanda Kim',         'Alexei Volkov',      '3+ Years',     '',                    '',          'High'),
        ('Regional Manager - APAC',        'Nikhil Sharma',       'Li Wei',             'Ready Now',    'Siti Rahayu',         '1-2 Years', 'Medium'),
        ('Regional Manager - EMEA',        'Hassan Al-Rashid',    'Giulia Moretti',     '1-2 Years',    'Andrei Popescu',      '3+ Years',  'Medium'),
        ('Regional Manager - Americas',    'Sophia Martinez',     'Tobias Müller',      '',             '',                    '',          ''),
        ('Head of Talent Acquisition',     'Cameron Wright',      'Amelia Thompson',    '1-2 Years',    '',                    '',          'Low'),
        ('Head of Learning & Development', 'Mei Xiong',           'Bashir Rahimi',      '3+ Years',     '',                    '',          'Low'),
        ('Head of Facilities & Admin',     'Jerome Deschamps',    'Oluwaseun Adebayo',  'Ready Now',    '',                    '',          'Low'),
        ('Head of Investor Relations',     'Claudia Weiss',       '',                   'No Successor', '',                    'N/A',       'Critical'),
    ]

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 24
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Succession Plan')
    print(f'  Rows: 1 header + 27 data rows (rows 2-28)')
    print(f'  No data validation dropdowns (to be added by agent)')
    print(f'  No conditional formatting (to be added by agent)')


create_initial()
