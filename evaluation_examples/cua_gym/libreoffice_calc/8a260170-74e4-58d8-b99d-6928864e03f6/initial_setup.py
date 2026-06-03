"""
Initial Setup: Process PDF receipts and update expense tracker
Task ID: osworld_multi_apps_doc_pdf_calc_004
Domain: libreoffice_calc (multi-app: PDF receipts + Calc expense log)

Creates:
  - /home/user/Desktop/receipts/coffee_shop.pdf  (receipt PDF)
  - /home/user/Desktop/receipts/uber_ride.pdf    (receipt PDF)
  - /home/user/Desktop/receipts/lunch.pdf        (receipt PDF)
  - /home/user/Desktop/expense_log.ods           (expense tracker with 5 existing rows, rows 7-9 EMPTY)
"""

import os
import shlex
import subprocess
import time

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_doc_pdf_calc_004'
DESKTOP = f'{WORKDIR}/Desktop'
RECEIPTS_DIR = f'{DESKTOP}/receipts'
EXPENSE_LOG = f'{DESKTOP}/expense_log.ods'
EXPENSE_LOG_XLSX = f'{DESKTOP}/expense_log.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def install_deps():
    """Install required Python packages on the VM."""
    subprocess.run(
        ['pip3', 'install', 'fpdf2', 'openpyxl', '--quiet'],
        capture_output=True,
    )


def create_pdf_receipt(filepath, receipt_title, date_str, vendor_name, amount_str, items):
    """Create a realistic PDF receipt."""
    from fpdf import FPDF

    pdf = FPDF()
    pdf.add_page()
    pdf.set_margins(20, 20, 20)

    # Header
    pdf.set_font('Helvetica', 'B', 18)
    pdf.cell(0, 12, receipt_title, ln=True, align='C')
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(0, 6, '123 Market Street, San Francisco, CA 94105', ln=True, align='C')
    pdf.cell(0, 6, 'Tel: (415) 555-0100  |  www.receipt.example.com', ln=True, align='C')
    pdf.ln(6)

    # Divider
    pdf.set_draw_color(100, 100, 100)
    pdf.line(20, pdf.get_y(), 190, pdf.get_y())
    pdf.ln(4)

    # Receipt details
    pdf.set_font('Helvetica', 'B', 11)
    pdf.cell(50, 7, 'RECEIPT', ln=False)
    pdf.set_font('Helvetica', '', 11)
    pdf.cell(0, 7, f'Date: {date_str}', ln=True, align='R')

    pdf.set_font('Helvetica', 'B', 11)
    pdf.cell(50, 7, 'Vendor:', ln=False)
    pdf.set_font('Helvetica', '', 11)
    pdf.cell(0, 7, vendor_name, ln=True)
    pdf.ln(4)

    # Items table header
    pdf.set_fill_color(230, 230, 230)
    pdf.set_font('Helvetica', 'B', 10)
    pdf.cell(100, 7, 'Description', border=1, fill=True, ln=False)
    pdf.cell(30, 7, 'Qty', border=1, fill=True, align='C', ln=False)
    pdf.cell(40, 7, 'Price', border=1, fill=True, align='R', ln=True)

    # Items
    pdf.set_font('Helvetica', '', 10)
    subtotal = 0.0
    for desc, qty, price in items:
        pdf.cell(100, 7, desc, border=1, ln=False)
        pdf.cell(30, 7, str(qty), border=1, align='C', ln=False)
        pdf.cell(40, 7, f'${price:.2f}', border=1, align='R', ln=True)
        subtotal += qty * price

    # Tax and total
    tax = round(subtotal * 0.0875, 2)
    total = round(subtotal + tax, 2)
    pdf.ln(4)
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(130, 7, '', ln=False)
    pdf.cell(20, 7, 'Subtotal:', ln=False, align='R')
    pdf.cell(20, 7, f'${subtotal:.2f}', ln=True, align='R')

    pdf.cell(130, 7, '', ln=False)
    pdf.cell(20, 7, 'Tax (8.75%):', ln=False, align='R')
    pdf.cell(20, 7, f'${tax:.2f}', ln=True, align='R')

    pdf.set_font('Helvetica', 'B', 11)
    pdf.cell(130, 8, '', ln=False)
    pdf.cell(20, 8, 'Amount:', ln=False, align='R')
    pdf.cell(20, 8, amount_str, ln=True, align='R')

    # Footer
    pdf.ln(8)
    pdf.set_font('Helvetica', 'I', 9)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 6, 'Thank you for your business!', ln=True, align='C')
    pdf.cell(0, 6, 'Please retain this receipt for your records.', ln=True, align='C')

    pdf.output(filepath)
    print(f'Created receipt: {filepath}')


def create_expense_log_xlsx():
    """Create the expense_log as xlsx first, then convert to ods."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Expenses'

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 12

    # Header row (row 1) — bold with light blue background
    headers = ['Date', 'Vendor', 'Amount']
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF', size=11)

    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 5 existing expense rows (rows 2-6) with realistic data
    existing_data = [
        ('2025-01-15', 'Office Depot',         42.30),
        ('2025-01-22', 'Delta Airlines',       385.00),
        ('2025-02-05', 'Marriott Hotels',      228.50),
        ('2025-02-18', 'FedEx Shipping',        17.85),
        ('2025-03-01', 'Adobe Creative Cloud',  54.99),  # row 6
    ]

    # Amount column format
    amount_format = '#,##0.00'
    date_format = 'yyyy-mm-dd'

    for r, (date_val, vendor, amount) in enumerate(existing_data, 2):
        ws.cell(row=r, column=1, value=date_val).number_format = date_format
        ws.cell(row=r, column=2, value=vendor)
        amt_cell = ws.cell(row=r, column=3, value=amount)
        amt_cell.number_format = amount_format

    # Rows 7-9 and row 10 are intentionally EMPTY (task will fill them)
    # (Do NOT pre-populate receipts data — that's the task!)

    wb.save(EXPENSE_LOG_XLSX)
    print(f'Created xlsx: {EXPENSE_LOG_XLSX}')


def convert_xlsx_to_ods():
    """Convert xlsx to ods using LibreOffice headless."""
    # Remove existing ods if any
    if os.path.exists(EXPENSE_LOG):
        os.remove(EXPENSE_LOG)

    result = subprocess.run(
        [
            'libreoffice', '--headless', '--convert-to', 'ods',
            '--outdir', DESKTOP,
            EXPENSE_LOG_XLSX,
        ],
        capture_output=True,
        text=True,
        timeout=60,
    )
    print('libreoffice convert stdout:', result.stdout)
    print('libreoffice convert stderr:', result.stderr)

    if os.path.exists(EXPENSE_LOG):
        print(f'Successfully converted to ODS: {EXPENSE_LOG}')
        # Remove xlsx intermediary
        os.remove(EXPENSE_LOG_XLSX)
    else:
        print(f'WARNING: ODS conversion failed, keeping xlsx as expense_log.xlsx')
        # Rename as fallback
        os.rename(EXPENSE_LOG_XLSX, f'{DESKTOP}/expense_log.xlsx')


def main():
    # Install dependencies
    install_deps()

    # Create receipts directory
    os.makedirs(RECEIPTS_DIR, exist_ok=True)
    print(f'Created directory: {RECEIPTS_DIR}')

    # Create Desktop directory if not exists
    os.makedirs(DESKTOP, exist_ok=True)

    # --- Create coffee_shop.pdf ---
    create_pdf_receipt(
        filepath=f'{RECEIPTS_DIR}/coffee_shop.pdf',
        receipt_title='Blue Bottle Coffee',
        date_str='2025-03-01',
        vendor_name='Blue Bottle Coffee',
        amount_str='$12.50',
        items=[
            ('Drip Coffee - Single Origin', 1, 5.00),
            ('Almond Croissant', 1, 4.75),
            ('Cold Brew (12 oz)', 1, 2.75),
        ],
    )

    # --- Create uber_ride.pdf ---
    create_pdf_receipt(
        filepath=f'{RECEIPTS_DIR}/uber_ride.pdf',
        receipt_title='Uber Technologies',
        date_str='2025-03-02',
        vendor_name='Uber Technologies',
        amount_str='$28.75',
        items=[
            ('UberX - SFO to Downtown', 1, 26.40),
            ('Service Fee', 1, 2.35),
        ],
    )

    # --- Create lunch.pdf ---
    create_pdf_receipt(
        filepath=f'{RECEIPTS_DIR}/lunch.pdf',
        receipt_title='The Sandwich Collective',
        date_str='2025-03-03',
        vendor_name='The Sandwich Collective',
        amount_str='$18.90',
        items=[
            ('Turkey Club Sandwich', 1, 12.50),
            ('House Salad', 1, 5.00),
            ('Sparkling Water', 1, 1.40),
        ],
    )

    # --- Create expense_log.ods ---
    create_expense_log_xlsx()
    convert_xlsx_to_ods()

    # Verify files exist
    print('\n--- Verification ---')
    for fpath in [
        f'{RECEIPTS_DIR}/coffee_shop.pdf',
        f'{RECEIPTS_DIR}/uber_ride.pdf',
        f'{RECEIPTS_DIR}/lunch.pdf',
        EXPENSE_LOG,
    ]:
        exists = os.path.exists(fpath)
        size = os.path.getsize(fpath) if exists else 0
        print(f'  {"OK" if exists else "MISSING"}: {fpath} ({size} bytes)')

    # --- GUI-ready startup ---
    # Open the expense_log file in LibreOffice Calc
    expense_file = EXPENSE_LOG if os.path.exists(EXPENSE_LOG) else f'{DESKTOP}/expense_log.xlsx'
    launch_gui(f'libreoffice --calc "{expense_file}"', delay_sec=2.0)

    # Open the receipts folder in Nautilus file manager
    launch_gui(f'nautilus "{RECEIPTS_DIR}"', delay_sec=1.0)

    print('\nGUI_READY: launched LibreOffice Calc + Nautilus with DISPLAY=:0')


main()
