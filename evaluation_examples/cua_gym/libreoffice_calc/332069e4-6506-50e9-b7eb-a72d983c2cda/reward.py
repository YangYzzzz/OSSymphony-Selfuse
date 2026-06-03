"""
Reward Script: Cross-sheet MAX formula on Dashboard
Task ID: calc_mcp_053
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): D2 contains a formula (not empty/literal)
  Component 2 (0.3): Formula uses MAX function referencing all 3 store sheets
  Component 3 (0.3): Formula references correct range B2:B100 on each store sheet
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_053'


def persist_app_state():
    """Save any unsaved LibreOffice edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(1.0)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Dashboard sheet must exist
    if 'Dashboard' not in wb.sheetnames:
        print("FAIL: 'Dashboard' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Dashboard']
    d2_value = ws.cell(row=2, column=4).value

    # Component 1: D2 contains a formula (0.4 points)
    # Initial state: D2 is None. Golden state: D2 has a formula starting with '='
    try:
        if d2_value is not None and isinstance(d2_value, str) and d2_value.startswith('='):
            print(f"PASS: Component 1 — D2 contains a formula: {d2_value} (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — D2 should contain a formula, found: {d2_value}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Formula uses MAX and references all 3 store sheets (0.3 points)
    # Normalize: remove spaces, uppercase for comparison
    # Accept both LibreOffice dot syntax (Store_A.B2) and Excel bang syntax (Store_A!B2)
    try:
        if d2_value and isinstance(d2_value, str):
            formula_upper = d2_value.upper().replace(" ", "")
            has_max = "MAX(" in formula_upper or "MAX (" in d2_value.upper()
            # Check all 3 store sheet references (accept . or ! as separator)
            has_store_a = bool(re.search(r'STORE_A[.!]', formula_upper))
            has_store_b = bool(re.search(r'STORE_B[.!]', formula_upper))
            has_store_c = bool(re.search(r'STORE_C[.!]', formula_upper))

            if has_max and has_store_a and has_store_b and has_store_c:
                print(f"PASS: Component 2 — MAX formula references Store_A, Store_B, Store_C (0.3 pts)")
                total_score += 0.3
            else:
                missing = []
                if not has_max:
                    missing.append("MAX function")
                if not has_store_a:
                    missing.append("Store_A ref")
                if not has_store_b:
                    missing.append("Store_B ref")
                if not has_store_c:
                    missing.append("Store_C ref")
                print(f"FAIL: Component 2 — Missing: {', '.join(missing)}. Formula: {d2_value}")
        else:
            print(f"FAIL: Component 2 — D2 is not a formula string")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Formula references correct range B2:B100 on each sheet (0.3 points)
    # The correct range is B2:B100 — check that each store reference includes this range
    try:
        if d2_value and isinstance(d2_value, str):
            formula_upper = d2_value.upper().replace(" ", "")
            # Match patterns like STORE_A.B2:B100 or STORE_A!B2:B100
            store_a_range = bool(re.search(r'STORE_A[.!]B2:B100', formula_upper))
            store_b_range = bool(re.search(r'STORE_B[.!]B2:B100', formula_upper))
            store_c_range = bool(re.search(r'STORE_C[.!]B2:B100', formula_upper))

            if store_a_range and store_b_range and store_c_range:
                print(f"PASS: Component 3 — All stores reference B2:B100 (0.3 pts)")
                total_score += 0.3
            else:
                missing = []
                if not store_a_range:
                    missing.append("Store_A.B2:B100")
                if not store_b_range:
                    missing.append("Store_B.B2:B100")
                if not store_c_range:
                    missing.append("Store_C.B2:B100")
                print(f"FAIL: Component 3 — Incorrect range for: {', '.join(missing)}. Formula: {d2_value}")
        else:
            print(f"FAIL: Component 3 — D2 is not a formula string")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state()

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
