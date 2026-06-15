"""
Initial Setup: Create cross-sheet formula task environment
Task ID: calc_mcp_053
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_053'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()

    # --- Dashboard sheet ---
    ws_dash = wb.active
    ws_dash.title = 'Dashboard'
    ws_dash['A1'] = 'Metric'
    ws_dash['B1'] = 'Description'
    ws_dash['C1'] = 'Category'
    ws_dash['D1'] = 'Value'

    ws_dash['A2'] = 'Overall Max Daily Sales'
    ws_dash['B2'] = 'Highest single-day sales across all stores'
    ws_dash['C2'] = 'Sales KPI'
    # D2 intentionally left EMPTY -- this is what the task asks the agent to fill

    ws_dash['A3'] = 'Report Date'
    ws_dash['B3'] = 'Date this dashboard was generated'
    ws_dash['C3'] = 'Admin'
    ws_dash['D3'] = '2026-03-31'

    ws_dash['A4'] = 'Stores Tracked'
    ws_dash['B4'] = 'Number of retail locations'
    ws_dash['C4'] = 'Admin'
    ws_dash['D4'] = 3

    # Style dashboard headers
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    for col in range(1, 5):
        cell = ws_dash.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws_dash.column_dimensions['A'].width = 25
    ws_dash.column_dimensions['B'].width = 42
    ws_dash.column_dimensions['C'].width = 14
    ws_dash.column_dimensions['D'].width = 16

    # --- Store sheets with daily sales data (B2:B100 = 99 days of data) ---
    store_configs = {
        'Store_A': {'base': 1200, 'var': 800, 'spike_day': 45, 'spike_val': 4850.75},
        'Store_B': {'base': 900, 'var': 600, 'spike_day': 72, 'spike_val': 5120.50},
        'Store_C': {'base': 1500, 'var': 1000, 'spike_day': 23, 'spike_val': 4975.25},
    }

    for store_name, cfg in store_configs.items():
        ws = wb.create_sheet(store_name)
        ws['A1'] = 'Day'
        ws['B1'] = 'Daily Sales'
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 14

        for day in range(1, 100):  # rows 2 through 100
            row = day + 1
            ws.cell(row=row, column=1, value=day)
            if day == cfg['spike_day']:
                sales = cfg['spike_val']
            else:
                sales = round(random.uniform(cfg['base'], cfg['base'] + cfg['var']), 2)
            ws.cell(row=row, column=2, value=sales)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
