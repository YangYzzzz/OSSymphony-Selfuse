"""
Initial Setup: Event log spreadsheet with raw date-time serial values in column C
Task ID: calc_fmt_numfmt_mixed_date_time_094
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_numfmt_mixed_date_time_094'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

# Convert a datetime to an Excel serial number
# Excel epoch: January 1, 1900 (serial 1), with the Lotus 1-2-3 bug (day 60 = Feb 29, 1900)
def datetime_to_serial(dt):
    epoch = datetime(1899, 12, 30)  # Excel's effective epoch
    delta = dt - epoch
    return delta.days + delta.seconds / 86400.0

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Event Log'

    # --- Headers (row 1) ---
    headers = ['Event ID', 'Type', 'Timestamp', 'User']
    bold_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold_font

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    # --- Event data ---
    event_types = [
        'LOGIN', 'LOGOUT', 'FILE_CREATE', 'FILE_DELETE', 'FILE_MODIFY',
        'PASSWORD_CHANGE', 'PERMISSION_CHANGE', 'CONFIG_UPDATE', 'BACKUP_START',
        'BACKUP_COMPLETE', 'ERROR', 'WARNING', 'ACCESS_DENIED', 'EXPORT', 'IMPORT'
    ]
    users = [
        'sarah.chen', 'marcus.johnson', 'emily.rodriguez', 'david.kim',
        'jessica.taylor', 'robert.nguyen', 'amanda.patel', 'christopher.lee',
        'michelle.wang', 'daniel.garcia', 'natalie.brown', 'james.wilson',
        'laura.martinez', 'kevin.thompson', 'ashley.white'
    ]

    # Generate 79 event log entries (rows 2-80)
    base_dt = datetime(2025, 3, 1, 8, 0, 0)
    rows = []
    event_id = 1001
    for i in range(79):
        # Spread events over ~30 days with realistic time gaps
        minutes_offset = i * 37 + (i % 7) * 113 + (i % 13) * 19
        event_dt = base_dt + timedelta(minutes=minutes_offset)

        serial = datetime_to_serial(event_dt)
        event_type = event_types[i % len(event_types)]
        user = users[i % len(users)]

        rows.append([f'EVT-{event_id}', event_type, serial, user])
        event_id += 1

    # Write data rows — column C uses 'General' format (raw serial numbers)
    for r, row_data in enumerate(rows, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            # Explicitly set column C to General format (raw serial numbers)
            if c == 3:
                cell.number_format = 'General'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Event Log')
    print(f'  Rows: 1 header + 79 data rows (rows 2-80)')
    print(f'  Column C: date-time serial values with General format')

create_initial()
