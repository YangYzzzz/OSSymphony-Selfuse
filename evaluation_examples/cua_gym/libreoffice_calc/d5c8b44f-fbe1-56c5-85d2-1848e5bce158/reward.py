"""
Reward Script: Format timestamp column (C2:C80) to display date and time as 'MM/DD/YYYY HH:MM'
Task ID: calc_fmt_numfmt_mixed_date_time_094
Domain: libreoffice_calc
Scoring:
  - Component 1: At least one cell in C2:C80 has the target format 'MM/DD/YYYY HH:MM' (0.4 pts)
  - Component 2: ALL 79 cells C2:C80 have exactly the format 'MM/DD/YYYY HH:MM' AND
                 no other data columns (A, B, D) have been modified (0.6 pts)
Total: 1.0
"""

import os
import openpyxl
from openpyxl.utils import get_column_letter

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmt_numfmt_mixed_date_time_094'
TARGET_FORMAT = 'MM/DD/YYYY HH:MM'
SHEET_NAME = 'Event Log'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Gate: verify the expected sheet exists
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Component 1: At least one cell in C2:C80 has the target number format (0.4 points)
    # This provides partial credit if the agent applied formatting to only some cells.
    # This FAILS on the initial file (all cells are 'General') and PASSES on the golden file.
    try:
        formatted_count = 0
        for row in range(2, 81):  # rows 2 to 80 inclusive (79 cells total)
            cell = ws.cell(row=row, column=3)
            if cell.number_format == TARGET_FORMAT:
                formatted_count += 1

        if formatted_count > 0:
            print(f"PASS: Component 1 — {formatted_count}/79 cells in C2:C80 have format "
                  f"'{TARGET_FORMAT}' (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — No cells in C2:C80 have format '{TARGET_FORMAT}'. "
                  f"Sample format found: '{ws.cell(row=2, column=3).number_format}'")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: ALL 79 cells in C2:C80 have exactly the target format AND
    #              columns A, B, D remain in 'General' format (no collateral changes) (0.6 points)
    # The task requires: (a) full range formatted, (b) no other cells modified.
    # This FAILS on the initial file (C2:C80 are 'General', not the target format)
    # and PASSES on the golden file.
    try:
        # Check all C2:C80 cells have the target format
        all_c_formatted = True
        wrong_c_examples = []
        for row in range(2, 81):
            cell = ws.cell(row=row, column=3)
            if cell.number_format != TARGET_FORMAT:
                all_c_formatted = False
                if len(wrong_c_examples) < 3:
                    wrong_c_examples.append(f"C{row}: '{cell.number_format}'")

        # Check that other data columns (A, B, D) are not inadvertently formatted
        other_cols_intact = True
        modified_other_examples = []
        for col in [1, 2, 4]:  # columns A, B, D
            col_letter = get_column_letter(col)
            for row in range(2, 81):
                cell = ws.cell(row=row, column=col)
                # These columns should remain in 'General' format (unchanged from initial)
                if cell.number_format not in ('General', '@', ''):
                    other_cols_intact = False
                    if len(modified_other_examples) < 3:
                        modified_other_examples.append(
                            f"{col_letter}{row}: '{cell.number_format}'"
                        )

        if all_c_formatted and other_cols_intact:
            print(f"PASS: Component 2 — All 79 cells C2:C80 have '{TARGET_FORMAT}' format "
                  f"and columns A, B, D are unmodified (0.6 pts)")
            total_score += 0.6
        elif not all_c_formatted:
            print(f"FAIL: Component 2 — Not all cells in C2:C80 are formatted correctly. "
                  f"Wrong format examples: {wrong_c_examples}")
        else:
            print(f"FAIL: Component 2 — Other columns were inadvertently modified: "
                  f"{modified_other_examples}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
