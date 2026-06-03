"""
Initial Setup: EventLog spreadsheet with log entries in column A, empty column B
Task ID: calc_fma_datevalue_text_073
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_fma_datevalue_text_073'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: EventLog ---
    ws = wb.active
    ws.title = 'EventLog'

    # Headers
    ws['A1'] = 'Log Entry'
    ws['B1'] = 'Event Date'
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)

    # Log data: 'YYYY-MM-DD: Event description' format (rows 2-12)
    log_entries = [
        '2024-01-15: Server restart',
        '2024-02-08: Database backup',
        '2024-03-22: Software update',
        '2024-04-03: Security patch',
        '2024-05-17: New user added',
        '2024-06-30: Quarterly audit',
        '2024-07-11: Config change',
        '2024-08-25: Maintenance window',
        '2024-09-09: Alert triggered',
        '2024-10-14: Report generated',
        '2024-11-28: Year-end review',
    ]

    for i, entry in enumerate(log_entries, start=2):
        ws.cell(row=i, column=1, value=entry)
        # Column B (Event Date) intentionally left empty for the agent to fill

    # Set column widths for readability
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

create_initial()
