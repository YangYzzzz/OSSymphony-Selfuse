"""
Reward Script: Extract date part from log entries and convert to date serial using DATEVALUE
Task ID: calc_fma_datevalue_text_073
Domain: libreoffice_calc
Scoring:
  Component 1: DATEVALUE(LEFT(...,10)) formulas in B2:B12 (0.5 points)
  Component 2: Date number format applied to B2:B12 (0.3 points)
  Component 3: All 11 cells have the formula — full coverage (0.2 points)
  Total: 1.0
"""

import os
import openpyxl
import re

WORKDIR = '/home/user'  # VM path — reward scripts run on the VM
TASK_ID = 'calc_fma_datevalue_text_073'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Task: Extract date part (first 10 chars) from column A log entries into column B
          using =DATEVALUE(LEFT(Ax,10)) formulas, formatted as dates.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'EventLog' sheet must exist
    if 'EventLog' not in wb.sheetnames:
        print("CRITICAL: 'EventLog' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['EventLog']

    # Component 1: At least one DATEVALUE(LEFT(...,10)) formula present in B2:B12 (0.5 points)
    # This checks the core transformation: extracting date part and converting to serial
    try:
        formula_cells_found = []
        for row in range(2, 13):  # B2:B12 (11 cells)
            val = ws.cell(row=row, column=2).value
            if val is not None and isinstance(val, str):
                # Match formula like =DATEVALUE(LEFT(A2,10)) — case insensitive
                normalized = val.strip().upper().replace(' ', '')
                if re.match(r'=DATEVALUE\(LEFT\(A\d+,10\)\)', normalized):
                    formula_cells_found.append(row)

        if len(formula_cells_found) >= 1:
            print(f"PASS: Component 1 — DATEVALUE(LEFT(...,10)) formulas found in {len(formula_cells_found)} cells: rows {formula_cells_found} (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — No DATEVALUE(LEFT(...,10)) formulas found in B2:B12")
            # Check what values are actually in column B
            for row in range(2, 13):
                val = ws.cell(row=row, column=2).value
                if val is not None:
                    print(f"  B{row} = {repr(val)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Date number format applied to cells with formulas in B2:B12 (0.3 points)
    # The cells should be formatted to display as dates (e.g., yyyy-mm-dd)
    try:
        date_format_cells = []
        for row in range(2, 13):
            cell = ws.cell(row=row, column=2)
            fmt = cell.number_format
            # Check for common date formats: yyyy-mm-dd, mm/dd/yyyy, etc.
            # Any format containing 'yy', 'mm', 'dd', or 'm', 'd' indicators is date-like
            if fmt and fmt != 'General' and fmt != '@' and fmt != '0':
                fmt_lower = fmt.lower()
                if any(indicator in fmt_lower for indicator in ['yyyy', 'yy', 'mm', 'dd', 'm/d', 'd/m']):
                    date_format_cells.append((row, fmt))

        if len(date_format_cells) >= 1:
            print(f"PASS: Component 2 — Date number format applied to {len(date_format_cells)} cells in B2:B12")
            for row, fmt in date_format_cells[:3]:  # show up to 3
                print(f"  B{row}: format = '{fmt}'")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — No date number format found in B2:B12")
            for row in range(2, 13):
                fmt = ws.cell(row=row, column=2).number_format
                if fmt and fmt != 'General':
                    print(f"  B{row}: format = '{fmt}'")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: All 11 cells B2:B12 have the DATEVALUE(LEFT(...,10)) formula (0.2 points)
    # This verifies complete coverage, not just partial
    try:
        all_formula_cells = []
        for row in range(2, 13):  # B2:B12
            val = ws.cell(row=row, column=2).value
            if val is not None and isinstance(val, str):
                normalized = val.strip().upper().replace(' ', '')
                if re.match(r'=DATEVALUE\(LEFT\(A\d+,10\)\)', normalized):
                    all_formula_cells.append(row)

        expected_rows = list(range(2, 13))  # rows 2 through 12
        missing_rows = [r for r in expected_rows if r not in all_formula_cells]

        if len(all_formula_cells) == 11 and not missing_rows:
            print(f"PASS: Component 3 — All 11 cells B2:B12 have DATEVALUE(LEFT(...,10)) formula (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — Only {len(all_formula_cells)}/11 cells have formula; missing rows: {missing_rows}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
