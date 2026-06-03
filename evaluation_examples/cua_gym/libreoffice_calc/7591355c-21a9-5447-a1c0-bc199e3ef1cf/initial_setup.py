"""
Initial Setup: Teacher Gradebook with multiple assessment types
Task ID: calc_wf_083
Domain: libreoffice_calc

Creates a gradebook with 25 students, 4 tests, 8 quizzes, 10 homework assignments,
and a participation score. Raw data only - no formulas, charts, or formatting.
"""

import os
import shlex
import subprocess
import time
import random
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_083'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(83)
    wb = openpyxl.Workbook()

    # --- Sheet: Gradebook ---
    ws = wb.active
    ws.title = 'Gradebook'

    # Headers
    headers = ['Student Name']
    headers += [f'Test {i}' for i in range(1, 5)]       # Test 1-4
    headers += [f'Quiz {i}' for i in range(1, 9)]       # Quiz 1-8
    headers += [f'HW {i}' for i in range(1, 11)]        # HW 1-10
    headers += ['Participation']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    # 25 students with realistic names
    students = [
        'Sarah Chen', 'Marcus Johnson', 'Emily Rodriguez', 'David Kim',
        'Jessica Patel', 'Andrew Thompson', 'Olivia Martinez', 'Ryan Nguyen',
        'Sophia Williams', 'Daniel Brown', 'Isabella Garcia', 'James Wilson',
        'Mia Anderson', 'Ethan Taylor', 'Ava Thomas', 'Noah Jackson',
        'Charlotte White', 'Liam Harris', 'Amelia Martin', 'Benjamin Lee',
        'Harper Robinson', 'Alexander Clark', 'Abigail Lewis', 'William Hall',
        'Ella Young'
    ]

    for r, name in enumerate(students, 2):
        ws.cell(row=r, column=1, value=name)
        col = 2
        # Tests (4): scores 55-100
        for _ in range(4):
            ws.cell(row=r, column=col, value=random.randint(55, 100))
            col += 1
        # Quizzes (8): scores 50-100
        for _ in range(8):
            ws.cell(row=r, column=col, value=random.randint(50, 100))
            col += 1
        # Homework (10): scores 40-100
        for _ in range(10):
            ws.cell(row=r, column=col, value=random.randint(40, 100))
            col += 1
        # Participation (1): score 60-100
        ws.cell(row=r, column=col, value=random.randint(60, 100))

    # Set reasonable column widths for readability
    ws.column_dimensions['A'].width = 22
    for c in range(2, len(headers) + 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(c)].width = 9

    # --- Sheet: Statistics (empty placeholder for task) ---
    wb.create_sheet('Statistics')

    # --- Sheet: Grade Distribution (empty placeholder for task) ---
    wb.create_sheet('Grade Distribution')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
