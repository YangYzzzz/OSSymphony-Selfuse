"""
Reward Script: Teacher Gradebook with Multiple Assessment Types
Task ID: calc_wf_083
Domain: libreoffice_calc

Scoring Rubric (6 components, total 1.0):
  1. Category average formulas with drop-lowest in Gradebook (Y-AA columns) — 0.25
  2. Weighted final grade formula in Gradebook (AB column) — 0.15
  3. Letter grade formula in Gradebook (AC column) — 0.10
  4. Statistics sheet populated with Mean/Median/StdDev/Min/Max — 0.20
  5. Grade Distribution sheet with counts + bar chart — 0.15
  6. Formatting: conditional formatting, bold headers, freeze panes — 0.15
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_083'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify required sheets exist
    required_sheets = ['Gradebook', 'Statistics', 'Grade Distribution']
    for sname in required_sheets:
        if sname not in wb.sheetnames:
            print(f"CRITICAL: Required sheet '{sname}' not found. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0

    ws = wb['Gradebook']

    # =========================================================================
    # Component 1: Category average formulas with drop-lowest (0.25 points)
    # Golden has: Y=AVERAGE(B:E), Z=(SUM(F:M)-SMALL(F:M,1))/7, AA=(SUM(N:W)-SMALL(N:W,1)-SMALL(N:W,2))/8
    # Initial has: Y1=None (no formulas)
    # =========================================================================
    try:
        comp1_score = 0.0
        # Check column Y header
        y_header = ws.cell(row=1, column=25).value  # col Y
        z_header = ws.cell(row=1, column=26).value  # col Z
        aa_header = ws.cell(row=1, column=27).value  # col AA

        has_calc_headers = (
            y_header is not None and
            z_header is not None and
            aa_header is not None
        )

        if not has_calc_headers:
            print("FAIL: Component 1 — Category average column headers not found (Y, Z, AA)")
        else:
            # Check formulas in row 2 for each category
            y2_val = ws.cell(row=2, column=25).value  # Test Avg
            z2_val = ws.cell(row=2, column=26).value  # Quiz Avg (drop lowest 1)
            aa2_val = ws.cell(row=2, column=27).value  # HW Avg (drop lowest 2)

            subpoints = 0.0

            # Test avg: should be AVERAGE of test cols (B-E)
            if y2_val and isinstance(y2_val, str) and 'AVERAGE' in y2_val.upper():
                subpoints += 1.0
                print(f"PASS: Test Avg formula found: {y2_val}")
            else:
                print(f"FAIL: Test Avg formula not found in Y2, got: {y2_val}")

            # Quiz avg: should use SUM and SMALL for drop-lowest
            if z2_val and isinstance(z2_val, str) and 'SUM' in z2_val.upper() and 'SMALL' in z2_val.upper():
                subpoints += 1.0
                print(f"PASS: Quiz Avg drop-lowest formula found: {z2_val}")
            else:
                print(f"FAIL: Quiz Avg drop-lowest formula not found in Z2, got: {z2_val}")

            # HW avg: should use SUM and SMALL for drop-lowest
            if aa2_val and isinstance(aa2_val, str) and 'SUM' in aa2_val.upper() and 'SMALL' in aa2_val.upper():
                subpoints += 1.0
                print(f"PASS: HW Avg drop-lowest formula found: {aa2_val}")
            else:
                print(f"FAIL: HW Avg drop-lowest formula not found in AA2, got: {aa2_val}")

            # Verify formulas are present for multiple students (not just row 2)
            formulas_present_count = 0
            for r in [2, 10, 20, 26]:
                if r > ws.max_row:
                    continue
                v = ws.cell(row=r, column=25).value
                if v and isinstance(v, str) and '=' in v:
                    formulas_present_count += 1
            if formulas_present_count >= 3:
                subpoints += 1.0
                print(f"PASS: Category avg formulas present across multiple rows ({formulas_present_count}/4 checked)")
            else:
                print(f"FAIL: Category avg formulas missing in some rows ({formulas_present_count}/4 checked)")

            comp1_score = (subpoints / 4.0) * 0.25
            total_score += comp1_score
            print(f"Component 1 score: {comp1_score:.4f}/0.25")

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Weighted final grade formula (0.15 points)
    # Golden has: AB2 = =Y2*0.4+Z2*0.25+AA2*0.2+X2*0.15
    # Initial has: AB2 = None
    # =========================================================================
    try:
        ab2_val = ws.cell(row=2, column=28).value  # col AB = Weighted Grade
        if ab2_val and isinstance(ab2_val, str):
            formula_upper = ab2_val.upper().replace(' ', '')
            # Check that formula references the weights 0.4, 0.25, 0.2, 0.15
            has_weights = (
                '0.4' in ab2_val and
                '0.25' in ab2_val and
                '0.2' in ab2_val and
                '0.15' in ab2_val
            )
            if has_weights:
                print(f"PASS: Component 2 — Weighted grade formula with correct weights: {ab2_val} (0.15 pts)")
                total_score += 0.15
            else:
                # Partial credit if there's any formula with multiplication
                if '*' in ab2_val:
                    print(f"PARTIAL: Component 2 — Weighted grade formula found but weights may differ: {ab2_val} (0.08 pts)")
                    total_score += 0.08
                else:
                    print(f"FAIL: Component 2 — Formula in AB2 doesn't appear to be weighted: {ab2_val}")
        else:
            print(f"FAIL: Component 2 — No weighted grade formula in AB2, got: {ab2_val}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Letter grade formula (0.10 points)
    # Golden has: AC2 = =IF(AB2>=90,"A",IF(AB2>=80,"B",IF(AB2>=70,"C",IF(AB2>=60,"D","F"))))
    # Initial has: AC2 = None
    # =========================================================================
    try:
        ac2_val = ws.cell(row=2, column=29).value  # col AC = Letter Grade
        if ac2_val and isinstance(ac2_val, str) and 'IF' in ac2_val.upper():
            # Should contain grade letters A, B, C, D, F
            val_upper = ac2_val.upper()
            has_grades = (
                '"A"' in ac2_val and
                '"B"' in ac2_val and
                '"F"' in ac2_val
            )
            if has_grades:
                print(f"PASS: Component 3 — Letter grade formula found: {ac2_val[:60]}... (0.10 pts)")
                total_score += 0.10
            else:
                print(f"PARTIAL: Component 3 — IF formula found but grade letters unclear: {ac2_val[:60]}... (0.05 pts)")
                total_score += 0.05
        else:
            print(f"FAIL: Component 3 — No letter grade formula in AC2, got: {ac2_val}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: Statistics sheet populated (0.20 points)
    # Golden has: Mean, Median, Std Dev, Min, Max rows with formulas
    # Initial has: empty Statistics sheet
    # =========================================================================
    try:
        ws_stats = wb['Statistics']
        comp4_score = 0.0

        # Check that Statistics has data (at least 5 data rows + header)
        if ws_stats.max_row >= 5 and ws_stats.max_column >= 2:
            # Check for statistical function keywords in cells
            stat_types_found = set()
            for r in range(1, ws_stats.max_row + 1):
                label = ws_stats.cell(row=r, column=1).value
                if label:
                    label_lower = str(label).lower()
                    if 'mean' in label_lower or 'average' in label_lower:
                        stat_types_found.add('mean')
                    elif 'median' in label_lower:
                        stat_types_found.add('median')
                    elif 'std' in label_lower or 'stdev' in label_lower or 'deviation' in label_lower:
                        stat_types_found.add('stddev')
                    elif 'min' in label_lower:
                        stat_types_found.add('min')
                    elif 'max' in label_lower:
                        stat_types_found.add('max')

            # Check formulas reference Gradebook
            formulas_with_ref = 0
            for r in range(2, ws_stats.max_row + 1):
                for c in range(2, ws_stats.max_column + 1):
                    v = ws_stats.cell(row=r, column=c).value
                    if v and isinstance(v, str) and 'Gradebook!' in v:
                        formulas_with_ref += 1

            # Score based on stat types found
            expected_stats = {'mean', 'median', 'stddev', 'min', 'max'}
            stats_ratio = len(stat_types_found & expected_stats) / len(expected_stats)
            comp4_score += stats_ratio * 0.10
            print(f"Stats types found: {stat_types_found} ({len(stat_types_found)}/5)")

            # Score based on cross-sheet formulas
            if formulas_with_ref >= 10:
                comp4_score += 0.10
                print(f"PASS: Statistics has {formulas_with_ref} cross-sheet formulas (0.10 pts)")
            elif formulas_with_ref >= 5:
                comp4_score += 0.05
                print(f"PARTIAL: Statistics has {formulas_with_ref} cross-sheet formulas (0.05 pts)")
            else:
                print(f"FAIL: Statistics has only {formulas_with_ref} cross-sheet formulas")

            total_score += comp4_score
            print(f"Component 4 score: {comp4_score:.4f}/0.20")
        else:
            print(f"FAIL: Component 4 — Statistics sheet is empty or too small ({ws_stats.max_row} rows, {ws_stats.max_column} cols)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # =========================================================================
    # Component 5: Grade Distribution sheet with counts + bar chart (0.15 points)
    # Golden has: A-F grade counts with COUNTIF/COUNTIFS formulas, 1 bar chart
    # Initial has: empty sheet, no chart
    # =========================================================================
    try:
        ws_dist = wb['Grade Distribution']
        comp5_score = 0.0

        # Check grade distribution data
        grade_labels_found = 0
        count_formulas_found = 0
        for r in range(2, ws_dist.max_row + 1):
            label = ws_dist.cell(row=r, column=1).value
            count_val = ws_dist.cell(row=r, column=2).value
            if label and isinstance(str(label), str):
                label_str = str(label).upper()
                if any(g in label_str for g in ['A', 'B', 'C', 'D', 'F']):
                    grade_labels_found += 1
            if count_val and isinstance(count_val, str) and 'COUNTIF' in count_val.upper():
                count_formulas_found += 1

        if grade_labels_found >= 4:
            comp5_score += 0.05
            print(f"PASS: Grade labels found: {grade_labels_found}")
        else:
            print(f"FAIL: Grade labels found: {grade_labels_found} (need >= 4)")

        if count_formulas_found >= 3:
            comp5_score += 0.05
            print(f"PASS: COUNTIF formulas found: {count_formulas_found}")
        else:
            print(f"FAIL: COUNTIF formulas found: {count_formulas_found} (need >= 3)")

        # Check for bar chart
        if len(ws_dist._charts) >= 1:
            chart = ws_dist._charts[0]
            chart_type = type(chart).__name__
            if 'Bar' in chart_type:
                comp5_score += 0.05
                print(f"PASS: Bar chart found on Grade Distribution sheet ({chart_type})")
            else:
                comp5_score += 0.03
                print(f"PARTIAL: Chart found but type is {chart_type}, expected BarChart")
        else:
            print(f"FAIL: No chart found on Grade Distribution sheet")

        total_score += comp5_score
        print(f"Component 5 score: {comp5_score:.4f}/0.15")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # =========================================================================
    # Component 6: Formatting — conditional formatting, bold headers, freeze panes (0.15 points)
    # Golden has: conditional formatting on AC2:AC26 and AB2:AB26, bold headers with blue fill, freeze A2
    # Initial has: no formatting, no freeze
    # =========================================================================
    try:
        comp6_score = 0.0

        # Check freeze panes (task asks to format for printing)
        if ws.freeze_panes is not None:
            comp6_score += 0.03
            print(f"PASS: Freeze panes set to {ws.freeze_panes}")
        else:
            print("FAIL: No freeze panes set")

        # Check bold headers in row 1
        bold_count = 0
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=c)
            if cell.font and cell.font.bold:
                bold_count += 1
        if bold_count >= 10:
            comp6_score += 0.03
            print(f"PASS: Bold headers found ({bold_count} bold cells in row 1)")
        else:
            print(f"FAIL: Bold headers — only {bold_count} bold cells in row 1")

        # Check header fill color (any non-default fill on header row)
        fill_count = 0
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=c)
            try:
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb and cell.fill.fgColor.rgb != '00000000':
                    fill_count += 1
            except:
                pass
        if fill_count >= 10:
            comp6_score += 0.03
            print(f"PASS: Header fill colors found ({fill_count} cells)")
        else:
            print(f"FAIL: Header fill — only {fill_count} cells with fill in row 1")

        # Check conditional formatting rules exist
        cf_rules = list(ws.conditional_formatting)
        if len(cf_rules) >= 1:
            # Count total rules across all ranges
            total_rules = sum(len(cf.rules) for cf in cf_rules)
            if total_rules >= 3:
                comp6_score += 0.06
                print(f"PASS: Conditional formatting — {total_rules} rules across {len(cf_rules)} ranges")
            elif total_rules >= 1:
                comp6_score += 0.03
                print(f"PARTIAL: Conditional formatting — {total_rules} rules (expected >= 3)")
            else:
                print("FAIL: Conditional formatting ranges exist but no rules")
        else:
            print("FAIL: No conditional formatting rules found")

        total_score += comp6_score
        print(f"Component 6 score: {comp6_score:.4f}/0.15")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Final score
    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.4f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook for LibreOffice
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
