"""
Reward Script: Parent-Teacher Conference Scheduling Sheet
Task ID: calc_grs_088
Domain: libreoffice_calc
Scoring:
  Component 1: 5 teacher sheets exist with correct structure (0.25)
  Component 2: Time slots are 15-min from 3pm-7pm with students & confirmation (0.20)
  Component 3: Color coding - booked=light blue, available=white, break=gray (0.15)
  Component 4: Master Schedule sheet consolidates all teachers (0.15)
  Component 5: Student Lookup sheet present with data (0.10)
  Component 6: Summary sheet with booking percentage per teacher (0.15)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_088'

# Expected teacher sheet names (task requires 5 teacher sheets)
EXPECTED_TEACHER_SHEETS = [
    'Hawkins_Math', 'Chen_ELA', 'Rivera_Science',
    'OBrien_SocStudies', 'Tanaka_ArtMusic'
]

# Expected 15-minute time slots from 3:00 PM to 6:45 PM
EXPECTED_TIME_SLOTS = [
    '3:00 PM', '3:15 PM', '3:30 PM', '3:45 PM',
    '4:00 PM', '4:15 PM', '4:30 PM', '4:45 PM',
    '5:00 PM', '5:15 PM', '5:30 PM', '5:45 PM',
    '6:00 PM', '6:15 PM', '6:30 PM', '6:45 PM'
]

# Color codes (ARGB)
LIGHT_BLUE = 'FFBDD7EE'
GRAY = 'FFC0C0C0'
WHITE = 'FFFFFFFF'


def get_fill_rgb(cell):
    """Safely get fill color RGB."""
    try:
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
            return str(cell.fill.fgColor.rgb)
    except Exception:
        pass
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    sheet_names = wb.sheetnames

    # =====================================================================
    # Component 1: 5 teacher sheets exist with correct grid structure (0.25)
    # Initial file has only 3 sheets (Conference Info, Teachers, Student Roster).
    # The task requires creating 5 teacher sheets each with a scheduling grid.
    # =====================================================================
    try:
        teacher_sheets_found = 0
        teacher_sheets_with_grid = 0

        for tname in EXPECTED_TEACHER_SHEETS:
            if tname in sheet_names:
                teacher_sheets_found += 1
                ws = wb[tname]
                # Verify grid structure: header row with Time Slot, and columns for
                # Tuesday/Thursday students + status + confirmed
                has_time_header = False
                has_status_cols = False
                for row in ws.iter_rows(min_row=1, max_row=10, max_col=7, values_only=False):
                    vals = [c.value for c in row]
                    if 'Time Slot' in vals:
                        has_time_header = True
                    # Check for status/confirmed columns
                    val_str = ' '.join(str(v) for v in vals if v)
                    if 'Status' in val_str and 'Confirmed' in val_str:
                        has_status_cols = True
                if has_time_header and has_status_cols:
                    teacher_sheets_with_grid += 1

        if teacher_sheets_with_grid >= 5:
            print(f"PASS: Component 1 -- All 5 teacher sheets with grid structure found (0.25 pts)")
            total_score += 0.25
        elif teacher_sheets_with_grid >= 3:
            partial = 0.15
            print(f"PARTIAL: Component 1 -- {teacher_sheets_with_grid}/5 teacher sheets with grid ({partial} pts)")
            total_score += partial
        elif teacher_sheets_found >= 1:
            partial = 0.05
            print(f"PARTIAL: Component 1 -- {teacher_sheets_found} teacher sheets found but incomplete grid ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 -- No teacher sheets found. Expected: {EXPECTED_TEACHER_SHEETS}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # =====================================================================
    # Component 2: Time slots are 15-min from 3pm-7pm with students assigned
    #              and confirmation column populated (0.20)
    # Initial file has no teacher sheets, so all of this is task-introduced.
    # =====================================================================
    try:
        sheets_with_valid_slots = 0
        sheets_with_students = 0
        sheets_with_confirmations = 0

        for tname in EXPECTED_TEACHER_SHEETS:
            if tname not in sheet_names:
                continue
            ws = wb[tname]

            # Collect time slot values from column A
            time_slots = []
            students_found = 0
            confirmations_found = 0

            for r in range(5, 25):  # rows where time data likely lives
                time_val = ws.cell(row=r, column=1).value
                if time_val and 'PM' in str(time_val):
                    time_slots.append(str(time_val).strip())

                # Check for student names in columns B or E (Tue/Thu)
                for col in [2, 5]:
                    student = ws.cell(row=r, column=col).value
                    if student and student != 'BREAK' and str(student).strip():
                        students_found += 1

                # Check confirmation columns D and G
                for col in [4, 7]:
                    conf = ws.cell(row=r, column=col).value
                    if conf and str(conf).strip() in ('Yes', 'Pending', 'No'):
                        confirmations_found += 1

            # Check that at least 16 time slots exist (3pm to 6:45pm = 16 slots)
            if len(time_slots) >= 16:
                sheets_with_valid_slots += 1
            if students_found >= 10:
                sheets_with_students += 1
            if confirmations_found >= 5:
                sheets_with_confirmations += 1

        # Need majority of teacher sheets to have correct slots + data
        if sheets_with_valid_slots >= 5 and sheets_with_students >= 5 and sheets_with_confirmations >= 5:
            print(f"PASS: Component 2 -- All 5 teachers have 16 time slots, students, and confirmations (0.20 pts)")
            total_score += 0.20
        elif sheets_with_valid_slots >= 3 and sheets_with_students >= 3:
            partial = 0.10
            print(f"PARTIAL: Component 2 -- {sheets_with_valid_slots}/5 valid slots, {sheets_with_students}/5 students ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- valid_slots={sheets_with_valid_slots}, students={sheets_with_students}, confirmations={sheets_with_confirmations}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # =====================================================================
    # Component 3: Color coding on teacher sheets (0.15)
    # Booked=light blue, Available=white, Break=gray
    # Initial has no teacher sheets so no colors to check.
    # =====================================================================
    try:
        sheets_with_colors = 0

        for tname in EXPECTED_TEACHER_SHEETS:
            if tname not in sheet_names:
                continue
            ws = wb[tname]

            booked_blue = 0
            break_gray = 0
            available_white = 0

            for r in range(5, 21):
                status_val = ws.cell(row=r, column=3).value  # Tue Status col
                cell_fill = get_fill_rgb(ws.cell(row=r, column=3))

                if status_val:
                    sv = str(status_val).strip().lower()
                    if sv == 'booked' and cell_fill == LIGHT_BLUE:
                        booked_blue += 1
                    elif sv == 'break' and cell_fill == GRAY:
                        break_gray += 1
                    elif sv == 'available' and cell_fill == WHITE:
                        available_white += 1

            # A teacher sheet should have at least some of each color
            if booked_blue >= 3 and break_gray >= 1:
                sheets_with_colors += 1

        if sheets_with_colors >= 5:
            print(f"PASS: Component 3 -- Color coding correct on all 5 teacher sheets (0.15 pts)")
            total_score += 0.15
        elif sheets_with_colors >= 3:
            partial = 0.08
            print(f"PARTIAL: Component 3 -- Color coding on {sheets_with_colors}/5 sheets ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 -- Only {sheets_with_colors}/5 sheets have correct color coding")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # =====================================================================
    # Component 4: Master Schedule sheet consolidates all teachers (0.15)
    # Must have all 5 teacher columns across 2 days with time slots.
    # Initial has no Master Schedule sheet.
    # =====================================================================
    try:
        master_found = False
        master_name = None
        for sn in sheet_names:
            if 'master' in sn.lower() and 'schedule' in sn.lower():
                master_found = True
                master_name = sn
                break

        if master_found:
            ws = wb[master_name]
            # Check for teacher names in header area
            header_text = ''
            for r in range(1, 6):
                for c in range(1, 15):
                    v = ws.cell(row=r, column=c).value
                    if v:
                        header_text += str(v) + ' '

            teachers_in_header = 0
            for t_keyword in ['Hawkins', 'Chen', 'Rivera', "O'Brien", 'Tanaka']:
                if t_keyword in header_text:
                    teachers_in_header += 1

            # Check that time slots are present in column A
            time_slots_count = 0
            for r in range(4, 25):
                v = ws.cell(row=r, column=1).value
                if v and 'PM' in str(v):
                    time_slots_count += 1

            # Check student data is populated in columns
            data_cells = 0
            for r in range(5, 21):
                for c in range(2, 12):
                    v = ws.cell(row=r, column=c).value
                    if v and str(v).strip() and str(v) != 'BREAK':
                        data_cells += 1

            if teachers_in_header >= 5 and time_slots_count >= 12 and data_cells >= 20:
                print(f"PASS: Component 4 -- Master Schedule has {teachers_in_header} teachers, {time_slots_count} time slots, {data_cells} data cells (0.15 pts)")
                total_score += 0.15
            elif teachers_in_header >= 3 and time_slots_count >= 8:
                partial = 0.08
                print(f"PARTIAL: Component 4 -- Master Schedule partial: {teachers_in_header} teachers, {time_slots_count} slots ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 4 -- Master Schedule incomplete: teachers={teachers_in_header}, slots={time_slots_count}, data={data_cells}")
        else:
            print(f"FAIL: Component 4 -- No 'Master Schedule' sheet found in {sheet_names}")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # =====================================================================
    # Component 5: Student Lookup sheet with conference data (0.10)
    # Shows parent their child's conference times across all teachers.
    # Initial has no Student Lookup sheet.
    # =====================================================================
    try:
        lookup_found = False
        lookup_name = None
        for sn in sheet_names:
            if 'lookup' in sn.lower() or ('student' in sn.lower() and sn != 'Student Roster'):
                lookup_found = True
                lookup_name = sn
                break

        if lookup_found:
            ws = wb[lookup_name]

            # Check for expected columns: Student Name, Teacher, Subject/Day/Time
            header_text = ''
            for c in range(1, 8):
                v = ws.cell(row=4, column=c).value
                if v:
                    header_text += str(v).lower() + ' '
            # Also check row 1 for alt header location
            for c in range(1, 8):
                v = ws.cell(row=1, column=c).value
                if v:
                    header_text += str(v).lower() + ' '

            has_student_col = 'student' in header_text
            has_teacher_col = 'teacher' in header_text

            # Count data rows (should have entries for multiple students)
            data_rows = 0
            for r in range(5, ws.max_row + 1):
                v = ws.cell(row=r, column=1).value
                if v and str(v).strip():
                    data_rows += 1

            if has_student_col and has_teacher_col and data_rows >= 20:
                print(f"PASS: Component 5 -- Student Lookup has headers and {data_rows} data rows (0.10 pts)")
                total_score += 0.10
            elif data_rows >= 5:
                partial = 0.05
                print(f"PARTIAL: Component 5 -- Student Lookup has {data_rows} rows ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 5 -- Student Lookup incomplete: student_col={has_student_col}, teacher_col={has_teacher_col}, rows={data_rows}")
        else:
            print(f"FAIL: Component 5 -- No Student Lookup sheet found in {sheet_names}")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    # =====================================================================
    # Component 6: Summary sheet with booking percentage per teacher (0.15)
    # Must show booking stats per teacher including booking %.
    # Initial has no Summary sheet.
    # =====================================================================
    try:
        summary_found = False
        summary_name = None
        for sn in sheet_names:
            if 'summary' in sn.lower():
                summary_found = True
                summary_name = sn
                break

        if summary_found:
            ws = wb[summary_name]

            # Check for Booking % or similar percentage column
            has_pct_header = False
            for r in range(1, 6):
                for c in range(1, 10):
                    v = ws.cell(row=r, column=c).value
                    if v and ('%' in str(v).lower() or 'booking' in str(v).lower() or 'percentage' in str(v).lower()):
                        has_pct_header = True

            # Check for teacher names and percentage values
            teachers_with_pct = 0
            for r in range(3, 20):
                teacher_val = ws.cell(row=r, column=1).value
                if teacher_val and any(kw in str(teacher_val) for kw in ['Hawkins', 'Chen', 'Rivera', "O'Brien", 'Tanaka']):
                    # Look for a percentage value in the row
                    for c in range(2, 10):
                        v = ws.cell(row=r, column=c).value
                        if v is not None:
                            try:
                                num = float(str(v).replace('%', ''))
                                if 0 < num <= 100:
                                    teachers_with_pct += 1
                                    break
                            except (ValueError, TypeError):
                                pass

            if has_pct_header and teachers_with_pct >= 5:
                print(f"PASS: Component 6 -- Summary has booking percentage for {teachers_with_pct} teachers (0.15 pts)")
                total_score += 0.15
            elif teachers_with_pct >= 3:
                partial = 0.08
                print(f"PARTIAL: Component 6 -- Summary has booking % for {teachers_with_pct}/5 teachers ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 6 -- Summary incomplete: pct_header={has_pct_header}, teachers_with_pct={teachers_with_pct}")
        else:
            print(f"FAIL: Component 6 -- No Summary sheet found in {sheet_names}")
    except Exception as e:
        print(f"ERROR: Component 6 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
