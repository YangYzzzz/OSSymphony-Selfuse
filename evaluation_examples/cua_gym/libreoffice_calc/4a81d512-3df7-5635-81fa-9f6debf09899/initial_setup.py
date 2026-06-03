"""
Initial Setup: School Parent-Teacher Conference Scheduling Sheet
Task ID: calc_grs_088
Domain: libreoffice_calc

Creates a workbook with raw student/teacher data that the agent will use
to build the complete conference scheduling system.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_088'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ---------------------------------------------------------------
    # Sheet 1: "Conference Info" - basic conference parameters
    # ---------------------------------------------------------------
    ws_info = wb.active
    ws_info.title = "Conference Info"

    ws_info["A1"] = "Parent-Teacher Conference Details"
    ws_info["A1"].font = Font(size=14, bold=True)

    ws_info["A3"] = "School:"
    ws_info["B3"] = "Maplewood Elementary School"
    ws_info["A4"] = "Conference Days:"
    ws_info["B4"] = "Tuesday, April 15 & Thursday, April 17"
    ws_info["A5"] = "Time Window:"
    ws_info["B5"] = "3:00 PM - 7:00 PM"
    ws_info["A6"] = "Slot Duration:"
    ws_info["B6"] = "15 minutes"
    ws_info["A7"] = "Number of Teachers:"
    ws_info["B7"] = 5
    ws_info["A8"] = "Students per Teacher:"
    ws_info["B8"] = 20

    ws_info["A10"] = "Color Coding Legend (to be applied):"
    ws_info["A11"] = "Booked"
    ws_info["B11"] = "Light Blue"
    ws_info["A12"] = "Available"
    ws_info["B12"] = "White"
    ws_info["A13"] = "Teacher Break"
    ws_info["B13"] = "Gray"

    for col in ['A', 'B']:
        ws_info.column_dimensions[col].width = 30

    # ---------------------------------------------------------------
    # Sheet 2: "Teachers" - list of teachers and their subjects
    # ---------------------------------------------------------------
    ws_teachers = wb.create_sheet("Teachers")

    teacher_headers = ["Teacher Name", "Subject", "Room", "Email"]
    for c, h in enumerate(teacher_headers, 1):
        cell = ws_teachers.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)

    teachers = [
        ["Mrs. Patricia Hawkins", "Mathematics", "Room 201", "p.hawkins@maplewood.edu"],
        ["Mr. David Chen", "English Language Arts", "Room 105", "d.chen@maplewood.edu"],
        ["Ms. Angela Rivera", "Science", "Room 310", "a.rivera@maplewood.edu"],
        ["Mr. James O'Brien", "Social Studies", "Room 208", "j.obrien@maplewood.edu"],
        ["Mrs. Keiko Tanaka", "Art & Music", "Room 115", "k.tanaka@maplewood.edu"],
    ]

    for r, row_data in enumerate(teachers, 2):
        for c, val in enumerate(row_data, 1):
            ws_teachers.cell(row=r, column=c, value=val)

    for col in ['A', 'B', 'C', 'D']:
        ws_teachers.column_dimensions[col].width = 28

    # ---------------------------------------------------------------
    # Sheet 3: "Student Roster" - all students with teacher assignments
    # ---------------------------------------------------------------
    ws_roster = wb.create_sheet("Student Roster")

    roster_headers = ["Student Name", "Grade", "Teacher", "Parent/Guardian", "Parent Email", "Parent Phone"]
    for c, h in enumerate(roster_headers, 1):
        cell = ws_roster.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)

    # 20 students per teacher = 100 total
    students_data = [
        # Mrs. Hawkins (Math) - 20 students
        ["Emma Thompson", "4th", "Mrs. Patricia Hawkins", "Laura Thompson", "l.thompson@email.com", "(555) 201-0001"],
        ["Liam Patel", "4th", "Mrs. Patricia Hawkins", "Priya Patel", "p.patel@email.com", "(555) 201-0002"],
        ["Sophia Martinez", "4th", "Mrs. Patricia Hawkins", "Carlos Martinez", "c.martinez@email.com", "(555) 201-0003"],
        ["Noah Williams", "4th", "Mrs. Patricia Hawkins", "Jessica Williams", "j.williams@email.com", "(555) 201-0004"],
        ["Olivia Brown", "4th", "Mrs. Patricia Hawkins", "Michael Brown", "m.brown@email.com", "(555) 201-0005"],
        ["Aiden Kim", "4th", "Mrs. Patricia Hawkins", "Soo-Jin Kim", "s.kim@email.com", "(555) 201-0006"],
        ["Isabella Garcia", "4th", "Mrs. Patricia Hawkins", "Maria Garcia", "m.garcia@email.com", "(555) 201-0007"],
        ["Mason Clark", "4th", "Mrs. Patricia Hawkins", "Robert Clark", "r.clark@email.com", "(555) 201-0008"],
        ["Mia Anderson", "4th", "Mrs. Patricia Hawkins", "Jennifer Anderson", "j.anderson@email.com", "(555) 201-0009"],
        ["Lucas Taylor", "4th", "Mrs. Patricia Hawkins", "Amanda Taylor", "a.taylor@email.com", "(555) 201-0010"],
        ["Charlotte Moore", "4th", "Mrs. Patricia Hawkins", "David Moore", "d.moore@email.com", "(555) 201-0011"],
        ["Ethan Jackson", "4th", "Mrs. Patricia Hawkins", "Stephanie Jackson", "s.jackson@email.com", "(555) 201-0012"],
        ["Amelia White", "4th", "Mrs. Patricia Hawkins", "Thomas White", "t.white@email.com", "(555) 201-0013"],
        ["James Lee", "4th", "Mrs. Patricia Hawkins", "Grace Lee", "g.lee@email.com", "(555) 201-0014"],
        ["Harper Davis", "4th", "Mrs. Patricia Hawkins", "Brian Davis", "b.davis@email.com", "(555) 201-0015"],
        ["Benjamin Wilson", "4th", "Mrs. Patricia Hawkins", "Karen Wilson", "k.wilson@email.com", "(555) 201-0016"],
        ["Evelyn Harris", "4th", "Mrs. Patricia Hawkins", "Kevin Harris", "k.harris@email.com", "(555) 201-0017"],
        ["Alexander Martin", "4th", "Mrs. Patricia Hawkins", "Linda Martin", "l.martin@email.com", "(555) 201-0018"],
        ["Abigail Robinson", "4th", "Mrs. Patricia Hawkins", "Daniel Robinson", "d.robinson@email.com", "(555) 201-0019"],
        ["Henry Walker", "4th", "Mrs. Patricia Hawkins", "Patricia Walker", "p.walker@email.com", "(555) 201-0020"],
        # Mr. Chen (ELA) - 20 students
        ["Ella Young", "3rd", "Mr. David Chen", "Susan Young", "s.young@email.com", "(555) 202-0001"],
        ["Jack Allen", "3rd", "Mr. David Chen", "Rachel Allen", "r.allen@email.com", "(555) 202-0002"],
        ["Scarlett King", "3rd", "Mr. David Chen", "Mark King", "m.king@email.com", "(555) 202-0003"],
        ["Owen Wright", "3rd", "Mr. David Chen", "Deborah Wright", "d.wright@email.com", "(555) 202-0004"],
        ["Grace Scott", "3rd", "Mr. David Chen", "Paul Scott", "p.scott@email.com", "(555) 202-0005"],
        ["Daniel Green", "3rd", "Mr. David Chen", "Lisa Green", "l.green@email.com", "(555) 202-0006"],
        ["Chloe Baker", "3rd", "Mr. David Chen", "Steven Baker", "s.baker@email.com", "(555) 202-0007"],
        ["Samuel Adams", "3rd", "Mr. David Chen", "Nancy Adams", "n.adams@email.com", "(555) 202-0008"],
        ["Zoey Nelson", "3rd", "Mr. David Chen", "George Nelson", "g.nelson@email.com", "(555) 202-0009"],
        ["Sebastian Hill", "3rd", "Mr. David Chen", "Barbara Hill", "b.hill@email.com", "(555) 202-0010"],
        ["Lily Ramirez", "3rd", "Mr. David Chen", "Jose Ramirez", "j.ramirez@email.com", "(555) 202-0011"],
        ["Matthew Campbell", "3rd", "Mr. David Chen", "Sharon Campbell", "s.campbell@email.com", "(555) 202-0012"],
        ["Aria Mitchell", "3rd", "Mr. David Chen", "Chris Mitchell", "c.mitchell@email.com", "(555) 202-0013"],
        ["Joseph Roberts", "3rd", "Mr. David Chen", "Helen Roberts", "h.roberts@email.com", "(555) 202-0014"],
        ["Riley Carter", "3rd", "Mr. David Chen", "Edward Carter", "e.carter@email.com", "(555) 202-0015"],
        ["David Phillips", "3rd", "Mr. David Chen", "Carol Phillips", "c.phillips@email.com", "(555) 202-0016"],
        ["Nora Evans", "3rd", "Mr. David Chen", "Timothy Evans", "t.evans@email.com", "(555) 202-0017"],
        ["Carter Turner", "3rd", "Mr. David Chen", "Donna Turner", "d.turner@email.com", "(555) 202-0018"],
        ["Hannah Parker", "3rd", "Mr. David Chen", "Ronald Parker", "r.parker@email.com", "(555) 202-0019"],
        ["Wyatt Collins", "3rd", "Mr. David Chen", "Dorothy Collins", "d.collins@email.com", "(555) 202-0020"],
        # Ms. Rivera (Science) - 20 students
        ["Penelope Edwards", "5th", "Ms. Angela Rivera", "Frank Edwards", "f.edwards@email.com", "(555) 203-0001"],
        ["Leo Stewart", "5th", "Ms. Angela Rivera", "Angela Stewart", "a.stewart@email.com", "(555) 203-0002"],
        ["Layla Sanchez", "5th", "Ms. Angela Rivera", "Miguel Sanchez", "m.sanchez@email.com", "(555) 203-0003"],
        ["Isaac Morris", "5th", "Ms. Angela Rivera", "Rebecca Morris", "r.morris@email.com", "(555) 203-0004"],
        ["Victoria Rogers", "5th", "Ms. Angela Rivera", "Andrew Rogers", "a.rogers@email.com", "(555) 203-0005"],
        ["Jayden Reed", "5th", "Ms. Angela Rivera", "Diane Reed", "d.reed@email.com", "(555) 203-0006"],
        ["Aurora Cook", "5th", "Ms. Angela Rivera", "Gary Cook", "g.cook@email.com", "(555) 203-0007"],
        ["Gabriel Morgan", "5th", "Ms. Angela Rivera", "Cynthia Morgan", "c.morgan@email.com", "(555) 203-0008"],
        ["Savannah Bell", "5th", "Ms. Angela Rivera", "Eric Bell", "e.bell@email.com", "(555) 203-0009"],
        ["Julian Murphy", "5th", "Ms. Angela Rivera", "Tammy Murphy", "t.murphy@email.com", "(555) 203-0010"],
        ["Bella Bailey", "5th", "Ms. Angela Rivera", "Dennis Bailey", "d.bailey@email.com", "(555) 203-0011"],
        ["Lincoln Rivera", "5th", "Ms. Angela Rivera", "Sandra Rivera", "s.rivera@email.com", "(555) 203-0012"],
        ["Claire Cooper", "5th", "Ms. Angela Rivera", "Jeffrey Cooper", "j.cooper@email.com", "(555) 203-0013"],
        ["Jaxon Richardson", "5th", "Ms. Angela Rivera", "Pamela Richardson", "p.richardson@email.com", "(555) 203-0014"],
        ["Lucy Cox", "5th", "Ms. Angela Rivera", "Wayne Cox", "w.cox@email.com", "(555) 203-0015"],
        ["Mateo Howard", "5th", "Ms. Angela Rivera", "Cheryl Howard", "c.howard@email.com", "(555) 203-0016"],
        ["Stella Ward", "5th", "Ms. Angela Rivera", "Randall Ward", "r.ward@email.com", "(555) 203-0017"],
        ["Anthony Torres", "5th", "Ms. Angela Rivera", "Janet Torres", "j.torres@email.com", "(555) 203-0018"],
        ["Hazel Peterson", "5th", "Ms. Angela Rivera", "Russell Peterson", "r.peterson@email.com", "(555) 203-0019"],
        ["Thomas Gray", "5th", "Ms. Angela Rivera", "Kathleen Gray", "k.gray@email.com", "(555) 203-0020"],
        # Mr. O'Brien (Social Studies) - 20 students
        ["Ellie James", "4th", "Mr. James O'Brien", "Philip James", "p.james@email.com", "(555) 204-0001"],
        ["Caleb Watson", "4th", "Mr. James O'Brien", "Brenda Watson", "b.watson@email.com", "(555) 204-0002"],
        ["Violet Brooks", "4th", "Mr. James O'Brien", "Roger Brooks", "r.brooks@email.com", "(555) 204-0003"],
        ["Ryan Kelly", "4th", "Mr. James O'Brien", "Heather Kelly", "h.kelly@email.com", "(555) 204-0004"],
        ["Natalie Sanders", "4th", "Mr. James O'Brien", "Terry Sanders", "t.sanders@email.com", "(555) 204-0005"],
        ["Nathan Price", "4th", "Mr. James O'Brien", "Virginia Price", "v.price@email.com", "(555) 204-0006"],
        ["Leah Bennett", "4th", "Mr. James O'Brien", "Joe Bennett", "j.bennett@email.com", "(555) 204-0007"],
        ["Christian Wood", "4th", "Mr. James O'Brien", "Marie Wood", "m.wood@email.com", "(555) 204-0008"],
        ["Aubrey Barnes", "4th", "Mr. James O'Brien", "Roy Barnes", "r.barnes@email.com", "(555) 204-0009"],
        ["Jonathan Ross", "4th", "Mr. James O'Brien", "Joyce Ross", "j.ross@email.com", "(555) 204-0010"],
        ["Addison Henderson", "4th", "Mr. James O'Brien", "Gerald Henderson", "g.henderson@email.com", "(555) 204-0011"],
        ["Nicholas Coleman", "4th", "Mr. James O'Brien", "Kathryn Coleman", "k.coleman@email.com", "(555) 204-0012"],
        ["Brooklyn Jenkins", "4th", "Mr. James O'Brien", "Albert Jenkins", "a.jenkins@email.com", "(555) 204-0013"],
        ["Andrew Perry", "4th", "Mr. James O'Brien", "Ruth Perry", "r.perry@email.com", "(555) 204-0014"],
        ["Skylar Powell", "4th", "Mr. James O'Brien", "Willie Powell", "w.powell@email.com", "(555) 204-0015"],
        ["Joshua Long", "4th", "Mr. James O'Brien", "Jean Long", "j.long@email.com", "(555) 204-0016"],
        ["Paisley Patterson", "4th", "Mr. James O'Brien", "Lawrence Patterson", "l.patterson@email.com", "(555) 204-0017"],
        ["Christopher Hughes", "4th", "Mr. James O'Brien", "Gloria Hughes", "g.hughes@email.com", "(555) 204-0018"],
        ["Madelyn Flores", "4th", "Mr. James O'Brien", "Arthur Flores", "a.flores@email.com", "(555) 204-0019"],
        ["Isaiah Washington", "4th", "Mr. James O'Brien", "Teresa Washington", "t.washington@email.com", "(555) 204-0020"],
        # Mrs. Tanaka (Art & Music) - 20 students
        ["Eleanor Butler", "3rd", "Mrs. Keiko Tanaka", "Wayne Butler", "w.butler@email.com", "(555) 205-0001"],
        ["Levi Simmons", "3rd", "Mrs. Keiko Tanaka", "Ann Simmons", "a.simmons@email.com", "(555) 205-0002"],
        ["Ruby Foster", "3rd", "Mrs. Keiko Tanaka", "Carl Foster", "c.foster@email.com", "(555) 205-0003"],
        ["Asher Gonzales", "3rd", "Mrs. Keiko Tanaka", "Rosa Gonzales", "r.gonzales@email.com", "(555) 205-0004"],
        ["Willow Bryant", "3rd", "Mrs. Keiko Tanaka", "Eugene Bryant", "e.bryant@email.com", "(555) 205-0005"],
        ["Ezra Alexander", "3rd", "Mrs. Keiko Tanaka", "Frances Alexander", "f.alexander@email.com", "(555) 205-0006"],
        ["Isla Russell", "3rd", "Mrs. Keiko Tanaka", "Lester Russell", "l.russell@email.com", "(555) 205-0007"],
        ["Carson Griffin", "3rd", "Mrs. Keiko Tanaka", "Irene Griffin", "i.griffin@email.com", "(555) 205-0008"],
        ["Luna Diaz", "3rd", "Mrs. Keiko Tanaka", "Mario Diaz", "m.diaz@email.com", "(555) 205-0009"],
        ["Miles Hayes", "3rd", "Mrs. Keiko Tanaka", "Norma Hayes", "n.hayes@email.com", "(555) 205-0010"],
        ["Emilia Myers", "3rd", "Mrs. Keiko Tanaka", "Fred Myers", "f.myers@email.com", "(555) 205-0011"],
        ["Dominic Ford", "3rd", "Mrs. Keiko Tanaka", "Evelyn Ford", "e.ford@email.com", "(555) 205-0012"],
        ["Ivy Hamilton", "3rd", "Mrs. Keiko Tanaka", "Stanley Hamilton", "s.hamilton@email.com", "(555) 205-0013"],
        ["Silas Graham", "3rd", "Mrs. Keiko Tanaka", "Lillian Graham", "l.graham@email.com", "(555) 205-0014"],
        ["Piper Sullivan", "3rd", "Mrs. Keiko Tanaka", "Ernest Sullivan", "e.sullivan@email.com", "(555) 205-0015"],
        ["Kai Wallace", "3rd", "Mrs. Keiko Tanaka", "Thelma Wallace", "t.wallace@email.com", "(555) 205-0016"],
        ["Jade West", "3rd", "Mrs. Keiko Tanaka", "Clarence West", "c.west@email.com", "(555) 205-0017"],
        ["Milo Cole", "3rd", "Mrs. Keiko Tanaka", "Edith Cole", "e.cole@email.com", "(555) 205-0018"],
        ["Daisy Hunt", "3rd", "Mrs. Keiko Tanaka", "Melvin Hunt", "m.hunt@email.com", "(555) 205-0019"],
        ["Felix Warren", "3rd", "Mrs. Keiko Tanaka", "Gladys Warren", "g.warren@email.com", "(555) 205-0020"],
    ]

    for r, row_data in enumerate(students_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_roster.cell(row=r, column=c, value=val)

    for col_letter, width in [('A', 22), ('B', 8), ('C', 24), ('D', 22), ('E', 28), ('F', 16)]:
        ws_roster.column_dimensions[col_letter].width = width

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
