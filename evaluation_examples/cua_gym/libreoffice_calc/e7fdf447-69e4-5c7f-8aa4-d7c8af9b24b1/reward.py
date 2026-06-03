"""
Reward Script: Place SUM formula for Q1+Q3 combined total in cell J6
Task ID: calc_fmb_sum_noncontiguous_006
Domain: libreoffice_calc
Scoring:
  Component 1: J6 contains a formula (not empty, not a literal value)        — 0.5 pts
  Component 2: Formula references both B6 and F6 (noncontiguous Q1+Q3)       — 0.3 pts
  Component 3: J6 formula present AND no other cells modified (compound)      — 0.2 pts
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmb_sum_noncontiguous_006'

# Known initial values that must NOT be changed
INITIAL_VALUES = {
    ('A1',): 'Annual Budget 2025',
    ('A2',): 'Department',
    ('B2',): 'Q1 Budget',
    ('C2',): 'Q1 Actual',
    ('D2',): 'Q2 Budget',
    ('E2',): 'Q2 Actual',
    ('F2',): 'Q3 Budget',
    ('G2',): 'Q3 Actual',
    ('H2',): 'Q4 Budget',
    ('I2',): 'Q4 Actual',
    ('J2',): 'Notes',
    ('A3',): 'Engineering',
    ('B3',): 58000,
    ('C3',): 54320,
    ('D3',): 65000,
    ('E3',): 68400,
    ('F3',): 55000,
    ('G3',): 52100,
    ('H3',): 68000,
    ('I3',): 71500,
    ('A4',): 'Marketing',
    ('B4',): 42000,
    ('C4',): 39800,
    ('D4',): 52000,
    ('E4',): 50300,
    ('F4',): 43500,
    ('G4',): 41900,
    ('H4',): 56200,
    ('I4',): 53700,
    ('A5',): 'Operations',
    ('B5',): 45000,
    ('C5',): 43600,
    ('D5',): 45000,
    ('E5',): 44800,
    ('F5',): 40000,
    ('G5',): 38900,
    ('H5',): 47000,
    ('I5',): 46300,
    ('J5',): 'Q1+Q3 Total',
    ('A6',): 'Total',
    ('B6',): 145000,
    ('D6',): 162000,
    ('F6',): 138500,
    ('H6',): 171200,
}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the sheet exists
    if 'Annual Budget' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Annual Budget' not found")
        print(f"Available sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Annual Budget']

    # Component 1: J6 contains a formula (not empty, not a literal value) — 0.5 points
    # This FAILS on initial (J6=None) and PASSES on golden (J6='=SUM(B6,F6)')
    try:
        j6_value = ws['J6'].value
        if j6_value is not None and isinstance(j6_value, str) and j6_value.strip().startswith('='):
            print(f"PASS: Component 1 — J6 contains a formula: {repr(j6_value)} (0.5 pts)")
            total_score += 0.5
        elif j6_value is None:
            print(f"FAIL: Component 1 — J6 is empty, expected a formula")
        else:
            print(f"FAIL: Component 1 — J6 contains non-formula value: {repr(j6_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not check J6: {e}")

    # Component 2: Formula in J6 references both B6 and F6 (noncontiguous Q1+Q3) — 0.3 points
    # This FAILS on initial (J6=None) and PASSES on golden (J6='=SUM(B6,F6)')
    try:
        j6_value = ws['J6'].value
        if j6_value is not None and isinstance(j6_value, str) and j6_value.strip().startswith('='):
            formula_upper = j6_value.upper().replace(' ', '')
            # Check that B6 and F6 are referenced (Q1 and Q3 totals respectively)
            has_b6 = bool(re.search(r'\bB6\b', formula_upper))
            has_f6 = bool(re.search(r'\bF6\b', formula_upper))
            # Check that Q2 (D6) and Q4 (H6) are NOT referenced (must be noncontiguous Q1+Q3 only)
            has_d6 = bool(re.search(r'\bD6\b', formula_upper))
            has_h6 = bool(re.search(r'\bH6\b', formula_upper))

            if has_b6 and has_f6 and not has_d6 and not has_h6:
                print(f"PASS: Component 2 — Formula references B6 (Q1) and F6 (Q3) without D6/H6: {repr(j6_value)} (0.3 pts)")
                total_score += 0.3
            elif has_b6 and has_f6 and (has_d6 or has_h6):
                print(f"FAIL: Component 2 — Formula includes Q2 or Q4 cells (D6 or H6): {repr(j6_value)}")
            elif not has_b6 and not has_f6:
                print(f"FAIL: Component 2 — Formula does not reference B6 or F6: {repr(j6_value)}")
            elif not has_b6:
                print(f"FAIL: Component 2 — Formula missing B6 (Q1 total): {repr(j6_value)}")
            elif not has_f6:
                print(f"FAIL: Component 2 — Formula missing F6 (Q3 total): {repr(j6_value)}")
        else:
            print(f"FAIL: Component 2 — J6 does not contain a formula, cannot check cell references")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not check formula cell references: {e}")

    # Component 3: J6 formula evaluates to the correct combined total AND no other cells modified — 0.2 points
    # This is a compound check: the formula must be present (J6 not None) AND the rest of the data is intact.
    # It FAILS on initial (J6=None, so the compound cannot hold) and PASSES on golden (J6 has formula + data intact).
    try:
        j6_value = ws['J6'].value
        if j6_value is None:
            # J6 is empty — cannot have correct total and intact data in the "completed task" sense
            print(f"FAIL: Component 3 — J6 is empty; formula result and data integrity check requires J6 to be set")
        elif not (isinstance(j6_value, str) and j6_value.strip().startswith('=')):
            print(f"FAIL: Component 3 — J6 is not a formula: {repr(j6_value)}")
        else:
            # J6 has a formula — now check that no other cells were modified
            modifications_found = []
            for (coord,), expected in INITIAL_VALUES.items():
                actual = ws[coord].value
                if actual != expected:
                    modifications_found.append(f"{coord}: expected {repr(expected)}, found {repr(actual)}")

            if not modifications_found:
                print(f"PASS: Component 3 — J6 formula present and no unintended cell modifications found (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 3 — Unexpected cell modifications detected (J6 has formula but other cells changed):")
                for m in modifications_found:
                    print(f"  - {m}")
    except Exception as e:
        print(f"ERROR: Component 3 — Could not check cell integrity: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
