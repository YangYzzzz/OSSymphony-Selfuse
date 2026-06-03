"""
Initial Setup: Annual Budget spreadsheet with Q1+Q3 noncontiguous sum task
Task ID: calc_fmb_sum_noncontiguous_006
Domain: libreoffice_calc

Creates a spreadsheet with quarterly budget data for 3 departments.
J6 is intentionally left empty — the agent must add =SUM(B6,F6) there.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_sum_noncontiguous_006'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Annual Budget'

    # --- Row 1: Title ---
    ws['A1'] = 'Annual Budget 2025'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:J1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # --- Row 2: Column Headers ---
    # Department column
    ws['A2'] = 'Department'
    ws['A2'].font = Font(bold=True)

    # Q1 headers (columns B-C)
    ws['B2'] = 'Q1 Budget'
    ws['B2'].font = Font(bold=True)
    ws['C2'] = 'Q1 Actual'
    ws['C2'].font = Font(bold=True)

    # Q2 headers (columns D-E)
    ws['D2'] = 'Q2 Budget'
    ws['D2'].font = Font(bold=True)
    ws['E2'] = 'Q2 Actual'
    ws['E2'].font = Font(bold=True)

    # Q3 headers (columns F-G)
    ws['F2'] = 'Q3 Budget'
    ws['F2'].font = Font(bold=True)
    ws['G2'] = 'Q3 Actual'
    ws['G2'].font = Font(bold=True)

    # Q4 headers (columns H-I)
    ws['H2'] = 'Q4 Budget'
    ws['H2'].font = Font(bold=True)
    ws['I2'] = 'Q4 Actual'
    ws['I2'].font = Font(bold=True)

    # J column header
    ws['J2'] = 'Notes'
    ws['J2'].font = Font(bold=True)

    # --- Rows 3-5: Department data ---
    departments = [
        # Dept,    Q1 Budget, Q1 Actual, Q2 Budget, Q2 Actual, Q3 Budget, Q3 Actual, Q4 Budget, Q4 Actual
        ['Engineering', 58000, 54320, 65000, 68400, 55000, 52100, 68000, 71500],
        ['Marketing',   42000, 39800, 52000, 50300, 43500, 41900, 56200, 53700],
        ['Operations',  45000, 43600, 45000, 44800, 40000, 38900, 47000, 46300],
    ]

    for r, row_data in enumerate(departments, 3):
        ws.cell(row=r, column=1, value=row_data[0])
        for c, val in enumerate(row_data[1:], 2):
            ws.cell(row=r, column=c, value=val)

    # --- Row 6: Quarter totals ---
    ws['A6'] = 'Total'
    ws['A6'].font = Font(bold=True)

    # Q1 total in B6
    ws['B6'] = 145000
    ws['B6'].font = Font(bold=True)

    # Q2 total in D6
    ws['D6'] = 162000
    ws['D6'].font = Font(bold=True)

    # Q3 total in F6
    ws['F6'] = 138500
    ws['F6'].font = Font(bold=True)

    # Q4 total in H6
    ws['H6'] = 171200
    ws['H6'].font = Font(bold=True)

    # --- Column J labels ---
    # J5: label for the target cell
    ws['J5'] = 'Q1+Q3 Total'
    ws['J5'].font = Font(bold=True)

    # J6 is intentionally LEFT EMPTY — this is what the agent must fill in

    # --- Column widths for readability ---
    ws.column_dimensions['A'].width = 16
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws.column_dimensions[col].width = 13

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('  Sheet: Annual Budget')
    print('  B6=145000 (Q1 Total), D6=162000 (Q2 Total)')
    print('  F6=138500 (Q3 Total), H6=171200 (Q4 Total)')
    print('  J5="Q1+Q3 Total", J6=<empty> (target cell)')


create_initial()
