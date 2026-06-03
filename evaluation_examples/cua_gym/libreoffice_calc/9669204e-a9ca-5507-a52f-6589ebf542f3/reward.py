"""
Reward Script: IFERROR/IF chained formula in LibreOffice Calc
Task ID: calc_fmb_iferror_chain_074
Domain: libreoffice_calc

Task: Enter the formula =IFERROR(IF(B2=0,0,A2/B2),"Error") in cell C2 of the
      'Unit Economics' sheet, so that #DIV/0! errors show 0 and #VALUE! errors
      show 'Error'.

Scoring:
  Component 1: C2 contains a formula that begins with IFERROR (0.4 pts)
               - FAILS on initial (C2 is empty) -> PASSES on golden
  Component 2: C2 formula contains nested IF(B2=0,...) for zero-division guard (0.3 pts)
               - FAILS on initial -> PASSES on golden
  Component 3: C2 formula references A2/B2 division AND the outer IFERROR uses "Error"
               as error value, AND no other data cells were modified (0.3 pts)
               - FAILS on initial -> PASSES on golden
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmb_iferror_chain_074'


def normalize_formula(formula):
    """Normalize formula: uppercase, remove spaces."""
    if not formula:
        return ''
    return str(formula).upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: check sheet 'Unit Economics' exists
    if 'Unit Economics' not in wb.sheetnames:
        print("FAIL: Sheet 'Unit Economics' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Unit Economics']

    # Component 1: C2 contains a formula that begins with =IFERROR (0.4 pts)
    # This checks that the agent entered a formula using IFERROR, which is the
    # outer wrapper required to catch #VALUE! errors.
    # On initial file, C2 is None (empty) -> FAILS
    # On golden file, C2 = '=IFERROR(IF(B2=0,0,A2/B2),"Error")' -> PASSES
    try:
        c2_value = ws['C2'].value
        if c2_value is not None and isinstance(c2_value, str):
            normalized = normalize_formula(c2_value)
            if normalized.startswith('=IFERROR('):
                print(f"PASS: Component 1 — C2 contains IFERROR formula: {repr(c2_value)} (0.4 pts)")
                total_score += 0.4
            else:
                print(f"FAIL: Component 1 — C2 has formula but does not start with =IFERROR: {repr(c2_value)}")
        else:
            print(f"FAIL: Component 1 — C2 is empty or not a formula string: {repr(c2_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — could not read C2: {e}")

    # Component 2: C2 formula contains nested IF(B2=0,...) for zero-division guard (0.3 pts)
    # This checks that the formula includes an IF condition to handle the case where
    # B2=0, returning 0 instead of causing #DIV/0!.
    # On initial file, C2 is None -> FAILS
    # On golden file, formula includes IF(B2=0,0,A2/B2) -> PASSES
    try:
        c2_value = ws['C2'].value
        if c2_value is not None and isinstance(c2_value, str):
            normalized = normalize_formula(c2_value)
            # Check for IF(B2=0 pattern
            if 'IF(B2=0' in normalized:
                print(f"PASS: Component 2 — C2 formula contains IF(B2=0,...) zero-division guard (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — C2 formula missing IF(B2=0,...) pattern: {repr(c2_value)}")
        else:
            print(f"FAIL: Component 2 — C2 is empty or not a formula: {repr(c2_value)}")
    except Exception as e:
        print(f"ERROR: Component 2 — could not read C2: {e}")

    # Component 3: C2 formula references A2/B2 division AND uses "Error" as error string,
    # AND no other cells (C3, C4) were accidentally modified (0.3 pts)
    # This checks that the division formula A2/B2 is inside the IF, that the IFERROR
    # fallback is the string "Error" (to show for #VALUE! errors), and that only C2
    # was changed (not C3 or C4, which should still be None).
    # On initial file, C2 is None -> the A2/B2 reference check fails -> FAILS
    # On golden file, formula has A2/B2 and "Error", and C3/C4 are still None -> PASSES
    try:
        c2_value = ws['C2'].value
        c3_value = ws['C3'].value
        c4_value = ws['C4'].value

        formula_ok = False
        if c2_value is not None and isinstance(c2_value, str):
            normalized = normalize_formula(c2_value)
            # Check for A2/B2 division reference and "ERROR" as string fallback
            has_division = 'A2/B2' in normalized
            has_error_string = '"ERROR"' in normalized
            formula_ok = has_division and has_error_string

        no_other_changes = (c3_value is None) and (c4_value is None)

        if formula_ok and no_other_changes:
            print(f"PASS: Component 3 — formula references A2/B2, uses 'Error' fallback, C3/C4 unchanged (0.3 pts)")
            total_score += 0.3
        elif not formula_ok:
            normalized_display = normalize_formula(c2_value) if c2_value else 'empty'
            print(f"FAIL: Component 3 — formula missing A2/B2 division or 'Error' string: {normalized_display}")
        else:
            print(f"FAIL: Component 3 — other cells were unexpectedly modified: C3={repr(c3_value)}, C4={repr(c4_value)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
