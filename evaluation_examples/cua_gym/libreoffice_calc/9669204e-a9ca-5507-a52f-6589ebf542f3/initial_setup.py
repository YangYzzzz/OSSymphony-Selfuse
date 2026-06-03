"""
Initial Setup: IFERROR chain formula for price-per-unit calculation
Task ID: calc_fmb_iferror_chain_074
Domain: libreoffice_calc
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'calc_fmb_iferror_chain_074'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Unit Economics ---
    ws = wb.active
    ws.title = 'Unit Economics'

    # Row 1: Headers
    ws['A1'] = 'Revenue'
    ws['B1'] = 'Units'
    ws['C1'] = 'Price per Unit'

    # Row 2: A2=0, B2=0 (both zero — will cause #DIV/0!). C2 is EMPTY (target cell).
    ws['A2'] = 0
    ws['B2'] = 0
    # C2 intentionally left empty — this is where the agent must enter the formula

    # Row 3: A3=15000, B3=150 (normal: 100 per unit)
    ws['A3'] = 15000
    ws['B3'] = 150

    # Row 4: A4='N/A', B4=50 (text in revenue — causes #VALUE! when dividing)
    ws['A4'] = 'N/A'
    ws['B4'] = 50

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheet: Unit Economics')
    print('  Row 1: Headers (Revenue, Units, Price per Unit)')
    print('  Row 2: A2=0, B2=0, C2=<empty> (target cell)')
    print('  Row 3: A3=15000, B3=150')
    print('  Row 4: A4=N/A, B4=50')


create_initial()
