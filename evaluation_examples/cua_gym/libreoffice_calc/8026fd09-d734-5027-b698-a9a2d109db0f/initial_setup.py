"""
Initial Setup: SUMIFS formula with mismatched criteria/sum ranges
Task ID: calc_tbl_061
Domain: libreoffice_calc

Creates a spreadsheet with sales data where the SUMIFS in E1 has
mismatched ranges: criteria range A2:A50 vs sum range D2:D100.
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_061'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SalesData"

    # --- Headers ---
    headers = ['Region', 'OrderAmount', 'Category', 'Value']
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Column widths ---
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 22

    # --- Data: 99 rows (rows 2 through 100) ---
    regions = ["East", "West", "North", "South"]
    categories = ["Electronics", "Furniture", "Clothing", "Food", "Office Supplies"]

    for r in range(2, 101):
        region = random.choice(regions)
        order_amount = round(random.uniform(200, 5000), 2)
        category = random.choice(categories)
        value = round(random.uniform(50, 3000), 2)

        ws.cell(row=r, column=1, value=region)
        ws.cell(row=r, column=2, value=order_amount)
        ws.cell(row=r, column=3, value=category)
        ws.cell(row=r, column=4, value=value)

    # --- BROKEN SUMIFS in E1: A range only goes to A50, not A100 ---
    ws["E1"] = '=SUMIFS(D2:D100,A2:A50,"East",B2:B100,">1000")'
    ws["E1"].font = Font(name="Calibri", size=11, bold=True)
    ws["E1"].number_format = '#,##0.00'

    # Label for E1
    ws["F1"] = "Conditional Sum (East, Order>1000)"
    ws["F1"].font = Font(name="Calibri", size=10, italic=True, color="666666")
    ws.column_dimensions["F"].width = 32

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
