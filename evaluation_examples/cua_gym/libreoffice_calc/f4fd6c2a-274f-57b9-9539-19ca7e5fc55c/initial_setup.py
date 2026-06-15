"""
Initial Setup: Create config_tracker.xlsx with Database and Settings sheets
Task ID: calc_gg5_040
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_040'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()

    # ========== Sheet 1: Database ==========
    ws_db = wb.active
    ws_db.title = 'Database'

    headers = ['ID', 'Category', 'Key', 'Value', 'LastModified']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws_db.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Categories
    categories = [
        'Network', 'Security', 'Storage', 'Display', 'Audio',
        'Power', 'Backup', 'Updates', 'Logging', 'Performance',
        'Authentication', 'Notifications'
    ]

    # Config key templates per category
    config_keys = {
        'Network': ['proxy_host', 'proxy_port', 'dns_primary', 'dns_secondary', 'gateway_ip',
                     'subnet_mask', 'dhcp_enabled', 'mtu_size', 'tcp_timeout', 'max_connections',
                     'bandwidth_limit', 'ipv6_enabled', 'firewall_zone', 'vlan_id', 'nic_bonding',
                     'nat_enabled', 'port_forwarding', 'ssl_verify', 'http_timeout', 'keep_alive_interval',
                     'retry_count', 'dns_cache_ttl', 'interface_priority', 'multicast_enabled', 'qos_policy'],
        'Security': ['encryption_algo', 'key_rotation_days', 'min_password_length', 'mfa_enabled',
                      'session_timeout_min', 'failed_login_limit', 'ip_whitelist', 'audit_log_enabled',
                      'certificate_path', 'tls_version', 'brute_force_lockout', 'password_complexity',
                      'token_expiry_hours', 'cors_policy', 'xss_protection', 'csrf_enabled',
                      'rate_limit_rpm', 'sandbox_mode', 'vulnerability_scan_freq', 'encryption_at_rest',
                      'data_masking', 'access_control_model', 'key_length_bits', 'hash_algorithm', 'secure_boot'],
        'Storage': ['max_disk_gb', 'raid_level', 'compression_enabled', 'dedup_enabled',
                     'snapshot_frequency', 'retention_days', 'io_scheduler', 'block_size_kb',
                     'cache_size_mb', 'thin_provisioning', 'replication_factor', 'mount_point',
                     'filesystem_type', 'quota_user_gb', 'encryption_enabled', 'tiering_policy',
                     'write_buffer_mb', 'read_ahead_kb', 'iops_limit', 'latency_threshold_ms',
                     'garbage_collection_freq', 'scrub_schedule', 'auto_expand', 'trim_enabled', 'zfs_pool_name'],
        'Display': ['resolution', 'refresh_rate_hz', 'color_depth', 'brightness_pct',
                     'gamma_value', 'night_mode_enabled', 'scaling_factor', 'vsync_enabled',
                     'hdr_enabled', 'color_profile', 'font_smoothing', 'dpi_override',
                     'multi_monitor_mode', 'screen_rotation', 'wallpaper_path',
                     'cursor_size', 'animation_speed', 'transparency_level', 'blue_light_filter', 'overscan_pct',
                     'aspect_ratio', 'panel_position', 'desktop_icons', 'screensaver_timeout', 'compositor_enabled'],
        'Audio': ['output_device', 'sample_rate_hz', 'bit_depth', 'channels',
                   'volume_master_pct', 'noise_cancellation', 'equalizer_preset', 'mic_gain_db',
                   'surround_enabled', 'latency_ms', 'buffer_size_frames', 'auto_mute_enabled',
                   'bluetooth_codec', 'spatial_audio', 'voice_enhancement', 'ducking_enabled',
                   'input_device', 'playback_speed', 'mono_mix', 'compressor_enabled',
                   'reverb_level', 'bass_boost_db', 'notification_sound', 'system_sounds_enabled', 'midi_device'],
        'Power': ['sleep_timeout_min', 'hibernate_enabled', 'cpu_governor', 'max_cpu_freq_ghz',
                   'turbo_boost_enabled', 'battery_threshold_pct', 'usb_suspend_enabled',
                   'wake_on_lan', 'ac_brightness_pct', 'battery_brightness_pct',
                   'thermal_policy', 'fan_curve_mode', 'power_button_action', 'lid_close_action',
                   'idle_timeout_min', 'gpu_power_mode', 'pcie_aspm_enabled', 'scheduled_shutdown',
                   'ups_enabled', 'energy_star_compliant', 'charge_limit_pct', 'fast_charge',
                   'adaptive_brightness', 'dimming_timeout_sec', 'suspend_to_disk'],
        'Backup': ['schedule_cron', 'destination_path', 'incremental_enabled', 'encryption_enabled',
                    'max_backup_count', 'compression_level', 'bandwidth_throttle_mbps', 'verify_after_backup',
                    'exclude_patterns', 'notification_email', 'retention_policy', 'parallel_streams',
                    'dedup_enabled', 'cloud_provider', 'bucket_name', 'region', 'lifecycle_days',
                    'versioning_enabled', 'cross_region_replication', 'restore_test_freq',
                    'snapshot_enabled', 'backup_window_start', 'backup_window_end', 'priority_level', 'log_retention_days'],
        'Updates': ['auto_update_enabled', 'check_frequency_hours', 'maintenance_window',
                     'rollback_enabled', 'staging_env', 'approval_required', 'max_concurrent_updates',
                     'reboot_policy', 'exclude_packages', 'critical_only', 'update_channel',
                     'delta_updates_enabled', 'bandwidth_limit_mbps', 'retry_attempts',
                     'notification_level', 'pre_update_snapshot', 'post_update_verify',
                     'update_log_path', 'proxy_for_updates', 'gpg_check_enabled',
                     'unattended_upgrades', 'schedule_day', 'schedule_hour', 'timeout_min', 'dry_run_first'],
        'Logging': ['log_level', 'max_file_size_mb', 'rotation_count', 'syslog_enabled',
                     'remote_server', 'format_pattern', 'timestamp_format', 'color_output',
                     'include_stacktrace', 'async_logging', 'buffer_size', 'flush_interval_sec',
                     'archive_enabled', 'archive_path', 'filter_pattern', 'json_format',
                     'correlation_id_enabled', 'sampling_rate', 'sensitive_data_masking',
                     'log_to_console', 'log_to_file', 'metrics_enabled', 'trace_enabled',
                     'audit_trail', 'retention_days'],
        'Performance': ['thread_pool_size', 'connection_pool_max', 'cache_ttl_sec', 'gc_frequency',
                         'memory_limit_mb', 'cpu_affinity', 'io_threads', 'prefetch_enabled',
                         'lazy_loading', 'batch_size', 'queue_depth', 'worker_count',
                         'timeout_sec', 'rate_limit', 'circuit_breaker_threshold', 'retry_backoff_ms',
                         'jit_enabled', 'profiling_enabled', 'heap_size_mb', 'stack_size_kb',
                         'numa_aware', 'huge_pages_enabled', 'swap_usage_limit', 'oom_score_adj', 'nice_level'],
        'Authentication': ['ldap_server', 'ldap_port', 'base_dn', 'bind_user',
                            'oauth_provider', 'token_lifetime_min', 'refresh_token_enabled',
                            'sso_enabled', 'saml_endpoint', 'jwt_algorithm', 'session_store',
                            'password_history_count', 'account_lockout_min', 'max_sessions',
                            'remember_me_days', 'two_factor_method', 'radius_server',
                            'kerberos_realm', 'certificate_auth_enabled', 'api_key_rotation_days',
                            'biometric_enabled', 'passkey_support', 'login_page_url', 'logout_redirect', 'idle_session_timeout'],
        'Notifications': ['email_enabled', 'smtp_server', 'smtp_port', 'smtp_tls',
                           'slack_webhook', 'teams_webhook', 'pagerduty_key', 'sns_topic_arn',
                           'push_enabled', 'digest_frequency', 'quiet_hours_start', 'quiet_hours_end',
                           'severity_threshold', 'template_path', 'max_retries', 'batch_enabled',
                           'dedup_window_min', 'escalation_policy', 'on_call_schedule',
                           'sms_enabled', 'sms_provider', 'voice_call_enabled', 'webhook_timeout_sec',
                           'custom_headers', 'rate_limit_per_min'],
    }

    # Value generators per key pattern
    def gen_value(key):
        if 'enabled' in key or 'compliant' in key or 'support' in key:
            return random.choice(['true', 'false'])
        if 'port' in key:
            return str(random.choice([22, 80, 443, 389, 636, 587, 993, 3306, 5432, 8080, 8443, 9090]))
        if 'timeout' in key or 'interval' in key or 'ttl' in key or 'lifetime' in key:
            return str(random.choice([5, 10, 15, 30, 60, 120, 300, 600, 900, 1800, 3600]))
        if 'size' in key or 'limit' in key or 'count' in key or 'max' in key or 'length' in key:
            return str(random.choice([4, 8, 16, 32, 64, 128, 256, 512, 1024, 2048, 4096]))
        if 'path' in key or 'url' in key or 'endpoint' in key:
            return f'/etc/app/{key.replace("_", "/")}'
        if 'pct' in key or 'level' in key or 'factor' in key or 'rate' in key:
            return str(random.choice([10, 20, 25, 30, 50, 60, 75, 80, 90, 95, 100]))
        if 'freq' in key or 'schedule' in key or 'cron' in key:
            return random.choice(['daily', 'weekly', 'hourly', '*/5 * * * *', '0 2 * * 0'])
        if 'ip' in key or 'server' in key or 'host' in key or 'gateway' in key:
            return f'192.168.{random.randint(1,254)}.{random.randint(1,254)}'
        if 'algo' in key or 'method' in key or 'type' in key or 'policy' in key or 'mode' in key:
            return random.choice(['AES-256', 'RSA-2048', 'SHA-256', 'HMAC', 'ECDSA', 'ChaCha20', 'default', 'adaptive', 'strict', 'balanced'])
        if 'email' in key:
            return random.choice(['admin@techcorp.com', 'ops@techcorp.com', 'alerts@techcorp.com'])
        if 'device' in key:
            return random.choice(['default', 'hdmi-0', 'usb-audio', 'bluetooth-speaker', 'analog-stereo'])
        if 'hz' in key:
            return str(random.choice([44100, 48000, 60, 75, 120, 144, 240]))
        if 'db' in key or 'gain' in key or 'boost' in key:
            return str(random.choice([-6, -3, 0, 3, 6, 9, 12]))
        if 'resolution' in key:
            return random.choice(['1920x1080', '2560x1440', '3840x2160'])
        if 'version' in key:
            return random.choice(['1.2', '1.3', 'TLSv1.3', '2.0', '3.1'])
        if 'day' in key or 'days' in key:
            return str(random.choice([1, 7, 14, 30, 60, 90, 180, 365]))
        if 'hour' in key:
            return str(random.choice([1, 2, 4, 6, 8, 12, 24]))
        if 'pattern' in key or 'format' in key:
            return random.choice(['%Y-%m-%d %H:%M:%S', 'ISO8601', 'RFC3339', '*.log', '*.tmp'])
        if 'name' in key or 'realm' in key or 'dn' in key:
            return random.choice(['techcorp.local', 'dc=techcorp,dc=com', 'TECHCORP.LOCAL', 'prod-cluster-01'])
        if 'webhook' in key:
            return 'https://hooks.techcorp.com/notify'
        return str(random.choice([1, 5, 10, 50, 100, 500, 'auto', 'default']))

    # Generate base date
    base_date = datetime(2025, 1, 1)

    # Generate 300 rows using round-robin across categories with varied keys
    data_rows = []
    row_id = 1
    # Distribute 300 entries roughly evenly (25 per category)
    for cat in categories:
        keys = config_keys[cat]
        for key in keys:
            mod_date = base_date + timedelta(days=random.randint(0, 400), hours=random.randint(0, 23), minutes=random.randint(0, 59))
            data_rows.append([
                f'CFG-{row_id:04d}',
                cat,
                key,
                gen_value(key),
                mod_date.strftime('%Y-%m-%d %H:%M'),
            ])
            row_id += 1

    # Shuffle for realism
    random.shuffle(data_rows)

    for r, row_data in enumerate(data_rows, 2):
        for c, val in enumerate(row_data, 1):
            ws_db.cell(row=r, column=c, value=val)

    # Set column widths
    ws_db.column_dimensions['A'].width = 12
    ws_db.column_dimensions['B'].width = 18
    ws_db.column_dimensions['C'].width = 30
    ws_db.column_dimensions['D'].width = 35
    ws_db.column_dimensions['E'].width = 20

    # Freeze header row
    ws_db.freeze_panes = 'A2'

    # ========== Sheet 2: Settings ==========
    ws_set = wb.create_sheet('Settings')

    # Title
    ws_set['A1'] = 'Configuration Lookup Tool'
    ws_set['A1'].font = Font(name='Calibri', size=14, bold=True)

    # Labels
    ws_set['A2'] = 'Select Category:'
    ws_set['A2'].font = Font(name='Calibri', size=11, bold=True)
    ws_set['A2'].alignment = Alignment(horizontal='right', vertical='center')

    # B2 dropdown with all categories
    cat_list = ','.join(sorted(categories))
    dv = DataValidation(
        type='list',
        formula1=f'"{cat_list}"',
        allow_blank=True,
        showDropDown=False,  # False = show the dropdown
    )
    dv.error = 'Please select a valid category'
    dv.errorTitle = 'Invalid Category'
    dv.prompt = 'Choose a configuration category'
    dv.promptTitle = 'Category'
    dv.add('B2')
    ws_set.add_data_validation(dv)

    # Set a default value in B2
    ws_set['B2'] = 'Network'
    ws_set['B2'].font = Font(name='Calibri', size=11)

    # Result headers in row 4
    result_headers = ['ID', 'Category', 'Key', 'Value', 'LastModified']
    for col, h in enumerate(result_headers, 3):  # C4, D4, E4, F4, G4
        cell = ws_set.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # C2:G2 and below are EMPTY (task requirement)
    # No formulas, no data in the result area

    # Column widths
    ws_set.column_dimensions['A'].width = 20
    ws_set.column_dimensions['B'].width = 20
    ws_set.column_dimensions['C'].width = 12
    ws_set.column_dimensions['D'].width = 18
    ws_set.column_dimensions['E'].width = 30
    ws_set.column_dimensions['F'].width = 35
    ws_set.column_dimensions['G'].width = 20

    # Instructions
    ws_set['A4'] = 'Instructions:'
    ws_set['A4'].font = Font(name='Calibri', size=10, bold=True)
    ws_set['A5'] = '1. Select a category from the dropdown in B2'
    ws_set['A6'] = '2. Results should appear in columns C-G'
    ws_set['A7'] = '3. The display updates automatically when category changes'
    for r in range(5, 8):
        ws_set.cell(row=r, column=1).font = Font(name='Calibri', size=9, color='666666')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
