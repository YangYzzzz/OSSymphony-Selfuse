"""
Reward Script: FILTER/IFERROR spill formula in Settings!C2 for config lookup
Task ID: calc_gg5_040
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20) - C2 contains a formula (not empty/literal)
  Component 2 (0.25) - Formula uses FILTER or INDEX/MATCH for lookup
  Component 3 (0.20) - Formula uses IFERROR for error handling
  Component 4 (0.15) - Formula references the Database sheet range
  Component 5 (0.20) - Formula references B2 (dropdown cell) as match criteria
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_040'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify that Settings!C2 contains a proper spill formula.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Settings sheet must exist
    if 'Settings' not in wb.sheetnames:
        print("CRITICAL: 'Settings' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Settings']
    c2_val = ws['C2'].value

    # Component 1: C2 contains a formula (0.20 points)
    # Initial state: C2 is None. Golden state: C2 has a formula string.
    try:
        if c2_val is not None and isinstance(c2_val, str) and c2_val.startswith('='):
            print(f"PASS: Component 1 — C2 contains a formula (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 — C2 is empty or not a formula. Found: {repr(c2_val)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # For remaining checks, we need C2 to be a formula string
    formula_upper = c2_val.upper().replace(" ", "") if isinstance(c2_val, str) else ""

    # Component 2: Formula uses FILTER or INDEX/MATCH for dynamic lookup (0.25 points)
    # Task requires XLOOKUP/INDEX/MATCH or FILTER with dynamic array spill
    try:
        has_filter = 'FILTER(' in formula_upper
        has_index_match = 'INDEX(' in formula_upper and 'MATCH(' in formula_upper
        has_xlookup = 'XLOOKUP(' in formula_upper
        if has_filter or has_index_match or has_xlookup:
            method = 'FILTER' if has_filter else ('INDEX/MATCH' if has_index_match else 'XLOOKUP')
            print(f"PASS: Component 2 — Formula uses {method} for lookup (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — No FILTER, INDEX/MATCH, or XLOOKUP found in formula")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Formula uses IFERROR wrapping (0.20 points)
    # Task explicitly requires IFERROR to handle no-match cases
    try:
        if 'IFERROR(' in formula_upper:
            print(f"PASS: Component 3 — Formula uses IFERROR wrapping (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 3 — No IFERROR found in formula")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Formula references the Database sheet (0.15 points)
    # The formula must pull data from the Database sheet
    try:
        if 'DATABASE!' in formula_upper or 'DATABASE!' in formula_upper.replace("'", ""):
            print(f"PASS: Component 4 — Formula references Database sheet (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — No reference to Database sheet found in formula")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Formula references B2 as the match criteria (0.20 points)
    # The dropdown value in B2 must be the lookup key
    try:
        # Check for B2 reference (could be $B$2, B2, Settings!B2, etc.)
        # We look for B2 pattern that is NOT part of Database!$B$2 range ref
        # The formula should reference B2 as the criteria value
        has_b2_ref = bool(re.search(r'(?:SETTINGS!\$?B\$?2|\$?B\$?2)', formula_upper))
        if has_b2_ref:
            print(f"PASS: Component 5 — Formula references B2 as match criteria (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 5 — No B2 reference found as match criteria")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
