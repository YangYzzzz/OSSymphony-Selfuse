"""
Reward Script: Sensor data analysis pipeline verification
Task ID: osworld_multi_apps_code_script_output_012
Domain: multi_apps (libreoffice_calc + os/scripts)
Scoring:
  Component 1: sensor_analysis.py script exists and contains core analysis logic (0.20 pts)
  Component 2: processed_sensor.csv has correct structure — hourly averages + boolean anomaly column (0.30 pts)
  Component 3: sensor_trends.png plot exists as a valid image file (0.20 pts)
  Component 4: processed_sensor.xlsx exists with conditional formatting highlighting anomaly rows red (0.30 pts)
Total: 1.0
"""

import os

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_code_script_output_012'

SCRIPT_PATH = f'{WORKDIR}/scripts/sensor_analysis.py'
PROCESSED_CSV_PATH = f'{WORKDIR}/data/processed_sensor.csv'
PLOT_PATH = f'{WORKDIR}/Desktop/sensor_trends.png'
PROCESSED_XLSX_PATH = f'{WORKDIR}/data/processed_sensor.xlsx'


def check_script_logic(script_path):
    """
    Return True if sensor_analysis.py contains all required core logic elements.
    """
    with open(script_path, 'r') as f:
        content = f.read()
    has_pandas = 'pandas' in content
    has_resample = 'resample' in content or '1h' in content or 'hourly' in content.lower()
    has_anomaly = ('std' in content) and ('anomal' in content.lower() or 'is_anomaly' in content)
    has_save = 'to_csv' in content or 'processed_sensor.csv' in content
    has_plot = 'matplotlib' in content or 'savefig' in content
    return sum([has_pandas, has_resample, has_anomaly, has_save, has_plot]) >= 4


def check_processed_csv(csv_path):
    """
    Return True if processed_sensor.csv has correct structure.
    """
    import pandas as pd
    df = pd.read_csv(csv_path)
    required = {'timestamp', 'temperature', 'humidity', 'sensor_id', 'is_anomaly'}
    has_cols = required.issubset(set(df.columns))
    has_3_sensors = df['sensor_id'].nunique() == 3
    unique_flags = set(df['is_anomaly'].unique())
    bool_vals = {True, False, 1, 0, 'True', 'False', 'true', 'false'}
    has_bool_col = unique_flags.issubset(bool_vals)
    has_hourly_rows = len(df) >= 400
    has_anomalies = df['is_anomaly'].astype(bool).sum() > 0
    return (has_cols, has_3_sensors, has_bool_col, has_hourly_rows, has_anomalies,
            len(df), df['sensor_id'].nunique(), int(df['is_anomaly'].astype(bool).sum()))


def check_plot_file(plot_path):
    """
    Return True if sensor_trends.png is a valid PNG image of substantial size.
    """
    with open(plot_path, 'rb') as f:
        magic = f.read(8)
    is_valid_png = (magic == b'\x89PNG\r\n\x1a\n')
    file_size = os.path.getsize(plot_path)
    return is_valid_png, file_size


def check_xlsx_cf(xlsx_path):
    """
    Return True if processed_sensor.xlsx has conditional formatting
    that highlights anomaly rows in red.
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    has_data = ws.max_row > 1
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    expected_headers = ['timestamp', 'temperature', 'humidity', 'sensor_id', 'is_anomaly']
    has_headers = all(h in headers for h in expected_headers)
    cf_rules = list(ws.conditional_formatting)
    has_any_cf = len(cf_rules) > 0
    has_red_fill = False
    cf_detail = "none"
    for cf in cf_rules:
        for rule in cf.rules:
            try:
                if rule.dxf is not None and rule.dxf.fill is not None:
                    for color_attr in ['fgColor', 'bgColor']:
                        try:
                            color_val = getattr(rule.dxf.fill, color_attr).rgb
                            if color_val and 'FF0000' in color_val.upper():
                                has_red_fill = True
                                cf_detail = f"range={cf}, {color_attr}={color_val}"
                        except Exception:
                            pass
            except Exception:
                pass
    return has_data, has_headers, has_any_cf, has_red_fill, cf_detail, len(cf_rules)


def verify_task():
    """
    Verify all task completion requirements with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # -------------------------------------------------------------------------
    # Component 1: sensor_analysis.py script exists and contains core logic
    # (0.20 points)
    # This FAILS on initial_env (script doesn't exist) and PASSES on golden_env
    # -------------------------------------------------------------------------
    try:
        script_exists = os.path.isfile(SCRIPT_PATH)
        if not script_exists:
            print(f"FAIL: Component 1 — sensor_analysis.py not found at {SCRIPT_PATH}")
        else:
            logic_ok = check_script_logic(SCRIPT_PATH)
            if logic_ok:
                print(f"PASS: Component 1 — sensor_analysis.py exists with core analysis logic (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 1 — sensor_analysis.py exists but missing key logic elements")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: processed_sensor.csv has hourly averages and boolean anomaly
    # (0.30 points)
    # This FAILS on initial_env (file doesn't exist) and PASSES on golden_env
    # -------------------------------------------------------------------------
    try:
        csv_exists = os.path.isfile(PROCESSED_CSV_PATH)
        if not csv_exists:
            print(f"FAIL: Component 2 — processed_sensor.csv not found at {PROCESSED_CSV_PATH}")
        else:
            (has_cols, has_3_sensors, has_bool_col, has_hourly_rows, has_anomalies,
             row_count, sensor_count, anomaly_count) = check_processed_csv(PROCESSED_CSV_PATH)
            csv_ok = has_cols and has_3_sensors and has_bool_col and has_hourly_rows and has_anomalies
            if csv_ok:
                total_score += 0.30
                print(f"PASS: Component 2 — processed_sensor.csv has correct structure: "
                      f"sensors={sensor_count}, rows={row_count}, "
                      f"bool_anomaly=True, anomaly_count={anomaly_count} (0.30 pts)")
            else:
                print(f"FAIL: Component 2 — processed_sensor.csv structure issues: "
                      f"has_cols={has_cols}, sensors={sensor_count}/3, "
                      f"bool_col={has_bool_col}, rows={row_count}(>=400), anomalies={has_anomalies}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: sensor_trends.png exists as a valid PNG image
    # (0.20 points)
    # This FAILS on initial_env (file doesn't exist) and PASSES on golden_env
    # -------------------------------------------------------------------------
    try:
        plot_exists = os.path.isfile(PLOT_PATH)
        if not plot_exists:
            print(f"FAIL: Component 3 — sensor_trends.png not found at {PLOT_PATH}")
        else:
            is_valid_png, file_size = check_plot_file(PLOT_PATH)
            plot_ok = is_valid_png and (file_size > 10000)
            if plot_ok:
                print(f"PASS: Component 3 — sensor_trends.png is a valid PNG ({file_size} bytes) (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 3 — sensor_trends.png: valid_png={is_valid_png}, "
                      f"size={file_size} bytes (need > 10000)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: processed_sensor.xlsx exists with conditional formatting
    # highlighting anomaly rows red
    # (0.30 points)
    # This FAILS on initial_env (file doesn't exist) and PASSES on golden_env
    # -------------------------------------------------------------------------
    try:
        xlsx_exists = os.path.isfile(PROCESSED_XLSX_PATH)
        if not xlsx_exists:
            print(f"FAIL: Component 4 — processed_sensor.xlsx not found at {PROCESSED_XLSX_PATH}")
        else:
            (has_data, has_headers, has_any_cf, has_red_fill,
             cf_detail, num_cf_rules) = check_xlsx_cf(PROCESSED_XLSX_PATH)
            # Accept if: has data, correct headers, AND (red fill confirmed OR has CF rules)
            xlsx_ok = has_data and has_headers and (has_red_fill or has_any_cf)
            if xlsx_ok:
                color_status = f"red_fill={has_red_fill}, {cf_detail}" if has_red_fill else f"cf_rules={num_cf_rules}"
                print(f"PASS: Component 4 — processed_sensor.xlsx with conditional formatting: "
                      f"{color_status} (0.30 pts)")
                total_score += 0.30
            else:
                print(f"FAIL: Component 4 — processed_sensor.xlsx: "
                      f"has_data={has_data}, has_headers={has_headers}, "
                      f"has_any_cf={has_any_cf}, has_red_fill={has_red_fill}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


verify_task()
