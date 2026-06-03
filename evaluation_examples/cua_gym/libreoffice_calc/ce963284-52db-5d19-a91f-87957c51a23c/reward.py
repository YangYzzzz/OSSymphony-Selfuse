"""
Reward Script: Use VLOOKUP to populate account manager column, sort by account manager,
and create a pivot table in Sheet2 with account managers as rows and quarterly totals as columns.
Task ID: osworld_calc_vlookup_pivot_combined_014
Domain: libreoffice_calc
Scoring:
  Component 1: VLOOKUP formulas in B2:B13 of Sheet1             (0.35 pts)
  Component 2: Sheet1 data sorted by Account Manager             (0.30 pts)
  Component 3: Pivot table in Sheet2 with correct structure/values (0.35 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_vlookup_pivot_combined_014'

# Lookup table: Client ID -> Account Manager (for sort verification)
LOOKUP_TABLE = {
    'C001': 'Alice Reynolds',
    'C002': 'Ben Carter',
    'C003': 'Clara Matthews',
    'C004': 'David Nguyen',
    'C005': 'Alice Reynolds',
    'C006': 'Ben Carter',
    'C007': 'Clara Matthews',
    'C008': 'David Nguyen',
    'C009': 'Alice Reynolds',
    'C010': 'Ben Carter',
    'C011': 'Clara Matthews',
    'C012': 'David Nguyen',
}

# Expected pivot table values (Account Manager -> [Q1, Q2, Q3, Q4])
EXPECTED_PIVOT = {
    'Alice Reynolds': [183300, 183300, 184200, 200800],
    'Ben Carter':     [177200, 179700, 189400, 186300],
    'Clara Matthews': [144500, 148400, 152300, 159100],
    'David Nguyen':   [169500, 170100, 170700, 185200],
}
EXPECTED_GRAND_TOTAL = [674500, 681500, 696600, 731400]


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Validate sheets exist (precondition gate)
    if 'Sheet1' not in wb.sheetnames or 'Sheet2' not in wb.sheetnames:
        print("CRITICAL: Sheet1 or Sheet2 not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws1 = wb['Sheet1']
    ws2 = wb['Sheet2']

    # -------------------------------------------------------------------------
    # Component 1: VLOOKUP formulas in B2:B13 of Sheet1 (0.35 points)
    # Initial file has None in B2:B13; golden has VLOOKUP formulas.
    # Verify that all 12 data rows in column B contain a VLOOKUP formula
    # referencing $H$2:$I$13.
    # -------------------------------------------------------------------------
    try:
        vlookup_count = 0
        vlookup_correct_ref_count = 0
        for row in range(2, 14):  # rows 2 through 13
            cell = ws1.cell(row=row, column=2)
            val = cell.value
            if val is not None and isinstance(val, str):
                val_upper = val.upper().replace(' ', '')
                if 'VLOOKUP' in val_upper:
                    vlookup_count += 1
                    # Check that the lookup range references $H$2:$I$13
                    if '$H$2:$I$13' in val.replace(' ', ''):
                        vlookup_correct_ref_count += 1

        if vlookup_count == 12 and vlookup_correct_ref_count == 12:
            print(f"PASS: Component 1 — All 12 rows in B2:B13 have VLOOKUP formulas referencing $H$2:$I$13 (0.35 pts)")
            total_score += 0.35
        elif vlookup_count == 12:
            # All VLOOKUP formulas exist but lookup range may vary
            print(f"PASS (partial): Component 1 — All 12 rows have VLOOKUP formulas but only {vlookup_correct_ref_count}/12 reference $H$2:$I$13. Awarding partial credit (0.20 pts)")
            total_score += 0.20
        elif vlookup_count > 0:
            print(f"FAIL: Component 1 — Only {vlookup_count}/12 rows in B2:B13 have VLOOKUP formulas")
        else:
            print(f"FAIL: Component 1 — No VLOOKUP formulas found in B2:B13 (all None or non-formula)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Sheet1 data sorted by Account Manager (0.30 points)
    # The sort groups rows by account manager: Alice Reynolds, Ben Carter,
    # Clara Matthews, David Nguyen (alphabetical). We verify that the
    # client IDs in column A follow this grouping.
    # Sort order expected: Alice clients (C001, C005, C009) first, then
    # Ben clients (C002, C006, C010), then Clara (C003, C007, C011),
    # then David (C004, C008, C012).
    # -------------------------------------------------------------------------
    try:
        # Derive the account manager for each row based on the lookup table
        managers_in_order = []
        for row in range(2, 14):
            client_id = ws1.cell(row=row, column=1).value
            if client_id in LOOKUP_TABLE:
                managers_in_order.append(LOOKUP_TABLE[client_id])
            else:
                managers_in_order.append(None)

        # Check the derived sequence is non-decreasing (alphabetical sort)
        manager_order = ['Alice Reynolds', 'Ben Carter', 'Clara Matthews', 'David Nguyen']
        sort_violations = 0
        prev_idx = -1
        for mgr in managers_in_order:
            if mgr is None:
                sort_violations += 1
                break
            idx = manager_order.index(mgr) if mgr in manager_order else -1
            if idx < prev_idx:
                sort_violations += 1
                break
            prev_idx = idx

        if sort_violations == 0 and None not in managers_in_order:
            print(f"PASS: Component 2 — Sheet1 rows are sorted by Account Manager (alphabetical order: {managers_in_order}) (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 2 — Sheet1 rows are NOT sorted by Account Manager. Derived order: {managers_in_order}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Pivot table in Sheet2 with correct structure and values (0.35 points)
    # Sheet2 must have:
    #   Row 1: headers — Account Manager, Q1 Sales, Q2 Sales, Q3 Sales, Q4 Sales
    #   Rows 2-5: one row per account manager with summed quarterly totals
    #   Row 6 (optional): Grand Total row
    # Expected values from context:
    #   Alice Reynolds: 183300, 183300, 184200, 200800
    #   Ben Carter:     177200, 179700, 189400, 186300
    #   Clara Matthews: 144500, 148400, 152300, 159100
    #   David Nguyen:   169500, 170100, 170700, 185200
    # -------------------------------------------------------------------------
    try:
        # Check headers
        header_row = [ws2.cell(row=1, column=c).value for c in range(1, 6)]
        expected_headers = ['Account Manager', 'Q1 Sales', 'Q2 Sales', 'Q3 Sales', 'Q4 Sales']
        headers_ok = (
            header_row[0] == 'Account Manager' and
            any('Q1' in str(h) for h in header_row[1:2]) and
            any('Q2' in str(h) for h in header_row[2:3]) and
            any('Q3' in str(h) for h in header_row[3:4]) and
            any('Q4' in str(h) for h in header_row[4:5])
        )

        # Collect data rows from Sheet2
        pivot_data = {}
        for row in range(2, ws2.max_row + 1):
            mgr = ws2.cell(row=row, column=1).value
            if mgr and str(mgr).strip() in EXPECTED_PIVOT:
                q1 = ws2.cell(row=row, column=2).value
                q2 = ws2.cell(row=row, column=3).value
                q3 = ws2.cell(row=row, column=4).value
                q4 = ws2.cell(row=row, column=5).value
                pivot_data[str(mgr).strip()] = [q1, q2, q3, q4]

        # Check if all 4 account managers are present with correct values
        pivot_correct_count = 0
        for mgr, expected_vals in EXPECTED_PIVOT.items():
            if mgr in pivot_data:
                actual_vals = pivot_data[mgr]
                vals_match = all(
                    actual_vals[i] is not None and abs(float(actual_vals[i]) - expected_vals[i]) <= 1.0
                    for i in range(4)
                )
                if vals_match:
                    pivot_correct_count += 1
                    print(f"  PASS pivot row: {mgr} = {actual_vals}")
                else:
                    print(f"  FAIL pivot row: {mgr} expected {expected_vals}, got {actual_vals}")
            else:
                print(f"  FAIL pivot row: {mgr} not found in Sheet2")

        if headers_ok and pivot_correct_count == 4:
            print(f"PASS: Component 3 — Pivot table in Sheet2 has correct headers and all 4 account manager rows with correct values (0.35 pts)")
            total_score += 0.35
        elif headers_ok and pivot_correct_count >= 2:
            partial = round(0.35 * pivot_correct_count / 4, 2)
            print(f"PASS (partial): Component 3 — Pivot table headers OK, {pivot_correct_count}/4 account manager rows correct. Awarding {partial} pts")
            total_score += partial
        elif pivot_correct_count > 0:
            partial = round(0.35 * pivot_correct_count / 4, 2)
            print(f"FAIL (partial): Component 3 — Headers missing or wrong, {pivot_correct_count}/4 account manager rows correct. Awarding {partial} pts")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Pivot table in Sheet2 is missing or has incorrect data. headers_ok={headers_ok}, rows found={len(pivot_data)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the VM environment
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
