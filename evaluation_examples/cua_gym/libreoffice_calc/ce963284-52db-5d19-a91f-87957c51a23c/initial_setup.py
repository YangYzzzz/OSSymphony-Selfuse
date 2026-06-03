"""
Initial Setup: VLOOKUP + Sort + Pivot Table task
Task ID: osworld_calc_vlookup_pivot_combined_014
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_vlookup_pivot_combined_014'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ── Sheet1: Sales Data ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Sheet1'

    # Headers in A1:F1
    headers = ['Client ID', 'Account Manager', 'Q1 Sales', 'Q2 Sales', 'Q3 Sales', 'Q4 Sales']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Lookup table header in H1:I1
    ws1.cell(row=1, column=8, value='Client ID').font = Font(bold=True)
    ws1.cell(row=1, column=9, value='Account Manager').font = Font(bold=True)

    # Sales data (Account Manager column is intentionally EMPTY — task is to fill it via VLOOKUP)
    # Unsorted order to make the sort meaningful
    sales_data = [
        # (Client ID, Q1, Q2, Q3, Q4)
        ('C003', 47200, 52100, 49800, 55300),
        ('C010', 38900, 41500, 43200, 40100),
        ('C007', 61500, 58900, 63400, 67200),
        ('C001', 85400, 79300, 88100, 92500),
        ('C012', 29700, 31200, 28900, 33400),
        ('C005', 54300, 57800, 51200, 59600),
        ('C002', 72100, 68400, 74800, 77300),
        ('C009', 43600, 46200, 44900, 48700),
        ('C011', 35800, 37400, 39100, 36600),
        ('C006', 66200, 69800, 71400, 68900),
        ('C004', 91300, 87600, 94200, 98700),
        ('C008', 48500, 51300, 47600, 53100),
    ]

    for r, (client_id, q1, q2, q3, q4) in enumerate(sales_data, 2):
        ws1.cell(row=r, column=1, value=client_id)
        # Column B (Account Manager) intentionally left EMPTY
        ws1.cell(row=r, column=3, value=q1)
        ws1.cell(row=r, column=4, value=q2)
        ws1.cell(row=r, column=5, value=q3)
        ws1.cell(row=r, column=6, value=q4)

    # Lookup table in columns H:I (rows 2-13)
    lookup_data = [
        ('C001', 'Alice Reynolds'),
        ('C002', 'Ben Carter'),
        ('C003', 'Clara Matthews'),
        ('C004', 'David Nguyen'),
        ('C005', 'Alice Reynolds'),
        ('C006', 'Ben Carter'),
        ('C007', 'Clara Matthews'),
        ('C008', 'David Nguyen'),
        ('C009', 'Alice Reynolds'),
        ('C010', 'Ben Carter'),
        ('C011', 'Clara Matthews'),
        ('C012', 'David Nguyen'),
    ]

    for r, (client_id, manager) in enumerate(lookup_data, 2):
        ws1.cell(row=r, column=8, value=client_id)
        ws1.cell(row=r, column=9, value=manager)

    # Column widths for readability
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 14
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 14
    ws1.column_dimensions['F'].width = 14
    ws1.column_dimensions['H'].width = 12
    ws1.column_dimensions['I'].width = 20

    # ── Sheet2: Empty (placeholder for pivot table) ─────────────────────
    ws2 = wb.create_sheet('Sheet2')
    ws2.cell(row=1, column=1, value='Pivot Table will be created here')
    ws2.cell(row=1, column=1).font = Font(italic=True, color='808080')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
