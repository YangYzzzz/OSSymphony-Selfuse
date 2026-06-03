"""
Reward Script: Social Media Content Calendar in LibreOffice Calc
Task ID: calc_grs_069
Domain: libreoffice_calc
Scoring:
  Component 1: Content Calendar sheet exists with correct structure (0.25)
  Component 2: Date grid covers 3 months with 5 platform columns (0.20)
  Component 3: Status columns (Caption Written, Graphics Ready, Scheduled, Published) with Y/N validation (0.15)
  Component 4: Campaign Tag dropdown (data validation) on column I (0.10)
  Component 5: Conditional formatting for content types with color coding (0.10)
  Component 6: Summary sheet with posts-per-platform-per-month and content type distribution (0.15)
  Component 7: Charts in Summary sheet (0.05)
"""

import os
import openpyxl
from datetime import datetime

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_069'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    sheet_names = wb.sheetnames

    # Component 1: Content Calendar sheet exists with correct column headers (0.25 points)
    try:
        # The golden file has a 'Content Calendar' sheet; initial has 'Raw Content Ideas' instead
        calendar_sheet = None
        for name in sheet_names:
            if 'calendar' in name.lower() or 'content calendar' in name.lower():
                calendar_sheet = wb[name]
                break

        if calendar_sheet is None:
            print("FAIL: Component 1 - No 'Content Calendar' sheet found")
        else:
            # Check required column headers
            headers = []
            for c in range(1, calendar_sheet.max_column + 1):
                v = calendar_sheet.cell(row=1, column=c).value
                if v:
                    headers.append(str(v).strip())

            required_headers = ['Date', 'Instagram', 'Facebook', 'Twitter/X', 'LinkedIn', 'TikTok']
            found_count = 0
            for rh in required_headers:
                if any(rh.lower() in h.lower() for h in headers):
                    found_count += 1

            if found_count >= 5:
                print(f"PASS: Component 1 - Content Calendar sheet found with {found_count}/{len(required_headers)} required headers (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 1 - Content Calendar found but only {found_count}/{len(required_headers)} required headers. Headers: {headers}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Date grid covers 3-month period with platform content (0.20 points)
    try:
        if calendar_sheet is None:
            print("FAIL: Component 2 - No Content Calendar sheet")
        else:
            # Check date column spans ~3 months (at least 60 data rows for ~90 days)
            dates = []
            for r in range(2, calendar_sheet.max_row + 1):
                v = calendar_sheet.cell(row=r, column=1).value
                if v is not None:
                    if isinstance(v, datetime):
                        dates.append(v)
                    elif isinstance(v, str):
                        try:
                            dates.append(datetime.strptime(v, '%Y-%m-%d'))
                        except ValueError:
                            pass

            if len(dates) < 60:
                print(f"FAIL: Component 2 - Only {len(dates)} dates found, expected ~90 for 3 months")
            else:
                # Check date range spans at least 2 months
                date_range_days = (max(dates) - min(dates)).days
                if date_range_days >= 59:
                    # Also verify platform columns have content (not all empty)
                    platform_cols = []
                    for c in range(1, calendar_sheet.max_column + 1):
                        h = calendar_sheet.cell(row=1, column=c).value
                        if h and str(h).strip() in ['Instagram', 'Facebook', 'Twitter/X', 'LinkedIn', 'TikTok']:
                            platform_cols.append(c)

                    has_content = False
                    for col in platform_cols:
                        for r in range(2, min(10, calendar_sheet.max_row + 1)):
                            if calendar_sheet.cell(row=r, column=col).value is not None:
                                has_content = True
                                break
                        if has_content:
                            break

                    if has_content:
                        print(f"PASS: Component 2 - {len(dates)} dates spanning {date_range_days} days with platform content (0.20 pts)")
                        total_score += 0.20
                    else:
                        print(f"FAIL: Component 2 - Dates found but platform columns are empty")
                else:
                    print(f"FAIL: Component 2 - Date range only {date_range_days} days, need >= 59")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Status columns with Y/N data validation (0.15 points)
    try:
        if calendar_sheet is None:
            print("FAIL: Component 3 - No Content Calendar sheet")
        else:
            # Check that status columns exist
            status_headers = ['Caption Written', 'Graphics Ready', 'Scheduled', 'Published']
            found_status = 0
            for c in range(1, calendar_sheet.max_column + 1):
                h = calendar_sheet.cell(row=1, column=c).value
                if h and any(sh.lower() in str(h).lower() for sh in status_headers):
                    found_status += 1

            # Check for Y/N data validation
            has_yn_validation = False
            if calendar_sheet.data_validations:
                for dv in calendar_sheet.data_validations.dataValidation:
                    if dv.type == 'list' and dv.formula1 and 'Y' in str(dv.formula1) and 'N' in str(dv.formula1):
                        has_yn_validation = True
                        break

            if found_status >= 3 and has_yn_validation:
                print(f"PASS: Component 3 - {found_status} status columns found with Y/N validation (0.15 pts)")
                total_score += 0.15
            elif found_status >= 3:
                print(f"PARTIAL: Component 3 - {found_status} status columns found but no Y/N validation (0.08 pts)")
                total_score += 0.08
            else:
                print(f"FAIL: Component 3 - Only {found_status} status columns found, need >= 3")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Campaign Tag dropdown (data validation on campaign column) (0.10 points)
    try:
        if calendar_sheet is None:
            print("FAIL: Component 4 - No Content Calendar sheet")
        else:
            # Find campaign column
            campaign_col = None
            for c in range(1, calendar_sheet.max_column + 1):
                h = calendar_sheet.cell(row=1, column=c).value
                if h and 'campaign' in str(h).lower():
                    campaign_col = c
                    break

            has_campaign_dv = False
            if campaign_col and calendar_sheet.data_validations:
                for dv in calendar_sheet.data_validations.dataValidation:
                    if dv.type == 'list':
                        # Check if sqref includes the campaign column
                        sqref_str = str(dv.sqref)
                        # Campaign col letter
                        from openpyxl.utils import get_column_letter
                        col_letter = get_column_letter(campaign_col)
                        if col_letter in sqref_str:
                            # Make sure it's not the Y/N one
                            if 'Y' not in str(dv.formula1) or len(str(dv.formula1)) > 10:
                                has_campaign_dv = True
                                break

            if campaign_col and has_campaign_dv:
                print(f"PASS: Component 4 - Campaign Tag column with dropdown validation found (0.10 pts)")
                total_score += 0.10
            elif campaign_col:
                print(f"PARTIAL: Component 4 - Campaign Tag column exists but no dropdown validation (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 4 - No Campaign Tag column found")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Conditional formatting for content types with color coding (0.10 points)
    try:
        if calendar_sheet is None:
            print("FAIL: Component 5 - No Content Calendar sheet")
        else:
            cf_rules = list(calendar_sheet.conditional_formatting)
            cf_count = 0
            for cf in cf_rules:
                cf_count += len(cf.rules)

            # Task requires 6 content type colors (Product=blue, Educational=green,
            # Behind Scenes=yellow, Promotions=orange, UGC=purple, Engagement=pink)
            if cf_count >= 4:
                print(f"PASS: Component 5 - {cf_count} conditional formatting rules found (0.10 pts)")
                total_score += 0.10
            elif cf_count >= 1:
                print(f"PARTIAL: Component 5 - Only {cf_count} CF rules, expected >= 4 (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 5 - No conditional formatting rules found")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    # Component 6: Summary sheet with platform-per-month and content type distribution (0.15 points)
    try:
        summary_sheet = None
        for name in sheet_names:
            if 'summary' in name.lower():
                summary_sheet = wb[name]
                break

        if summary_sheet is None:
            print("FAIL: Component 6 - No Summary sheet found")
        else:
            # Check for platform names in the summary
            platforms_found = set()
            content_types_found = set()
            expected_platforms = {'instagram', 'facebook', 'twitter/x', 'linkedin', 'tiktok'}
            expected_content_types = {'product posts', 'educational content', 'behind the scenes',
                                     'promotions', 'user generated content', 'engagement questions'}

            for row in summary_sheet.iter_rows(min_row=1, max_row=summary_sheet.max_row, max_col=summary_sheet.max_column, values_only=False):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        val_lower = cell.value.strip().lower()
                        if val_lower in expected_platforms:
                            platforms_found.add(val_lower)
                        if val_lower in expected_content_types:
                            content_types_found.add(val_lower)

            has_platforms = len(platforms_found) >= 4
            has_content_types = len(content_types_found) >= 4

            if has_platforms and has_content_types:
                print(f"PASS: Component 6 - Summary has {len(platforms_found)} platforms and {len(content_types_found)} content types (0.15 pts)")
                total_score += 0.15
            elif has_platforms or has_content_types:
                print(f"PARTIAL: Component 6 - Platforms: {len(platforms_found)}, Content types: {len(content_types_found)} (0.08 pts)")
                total_score += 0.08
            else:
                print(f"FAIL: Component 6 - Summary lacks platform/content type data")
    except Exception as e:
        print(f"ERROR: Component 6 - {e}")

    # Component 7: Charts in Summary sheet (0.05 points)
    try:
        if summary_sheet is None:
            print("FAIL: Component 7 - No Summary sheet")
        else:
            chart_count = len(summary_sheet._charts)
            if chart_count >= 1:
                print(f"PASS: Component 7 - {chart_count} chart(s) in Summary sheet (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 7 - No charts in Summary sheet")
    except Exception as e:
        print(f"ERROR: Component 7 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
