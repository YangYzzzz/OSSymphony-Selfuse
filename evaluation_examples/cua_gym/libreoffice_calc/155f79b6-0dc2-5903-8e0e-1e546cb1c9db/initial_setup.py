"""
Initial Setup: Social media content calendar for 3-month period
Task ID: calc_grs_069
Domain: libreoffice_calc

Creates a workbook with raw content data that needs to be organized into
a proper content calendar with conditional formatting, status tracking,
and summary statistics.
"""

import os
import shlex
import subprocess
import time
import random
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_069'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Raw Content Ideas ---
    ws1 = wb.active
    ws1.title = 'Raw Content Ideas'

    headers = ['Date', 'Platform', 'Content Type', 'Post Topic/Theme', 'Campaign']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)
        ws1.cell(row=1, column=col).font = Font(bold=True)

    platforms = ['Instagram', 'Facebook', 'Twitter/X', 'LinkedIn', 'TikTok']
    content_types = [
        'Product Posts', 'Educational Content', 'Behind the Scenes',
        'Promotions', 'User Generated Content', 'Engagement Questions'
    ]
    campaigns = ['Spring Launch', 'Earth Day', 'Q1 Push', 'Brand Awareness', '']

    topics = {
        'Product Posts': [
            'New spring collection showcase', 'Product feature highlight - durability test',
            'Customer favorite product roundup', 'Limited edition reveal',
            'Product comparison guide', 'Bestseller of the month',
            'New color launch preview', 'Product in action - outdoor shoot',
            'Ingredient spotlight - organic sourcing', 'Bundle deal showcase',
        ],
        'Educational Content': [
            'How to style our products for work', '5 tips for sustainable living',
            'Industry trends 2025 breakdown', 'Material care guide',
            'Behind the design process explained', 'Sustainability report highlights',
            'How we source our materials', 'Guide to choosing the right size',
            'Benefits of eco-friendly packaging', 'FAQ answered by our founder',
        ],
        'Behind the Scenes': [
            'Day in the life at HQ', 'Warehouse tour and packing process',
            'Team brainstorm session recap', 'Photo shoot preparation',
            'Meet the team - Engineering dept', 'Office pets Friday',
            'Factory visit vlog', 'Design team mood board creation',
            'Quality control walkthrough', 'Meet our new intern',
        ],
        'Promotions': [
            'Flash sale announcement - 24hr only', 'Free shipping weekend promo',
            'Buy 2 get 1 free campaign', 'Loyalty rewards double points',
            'Early access for subscribers', 'Seasonal clearance event',
            'Referral bonus program launch', 'Birthday month discount',
            'Newsletter exclusive 15% off', 'VIP member appreciation week',
        ],
        'User Generated Content': [
            'Customer review spotlight - @sarah_adventures', 'Repost: Creative unboxing by @designjay',
            'Customer photo contest winner', 'Testimonial video - Marcus J.',
            'Fan art feature Friday', 'Customer transformation story',
            'User styling challenge entries', 'Community poll results showcase',
            'Best customer photos this month', 'Influencer collab - @eco_emma',
        ],
        'Engagement Questions': [
            'Which color do you prefer? Poll', 'Caption this photo contest',
            'What product should we bring back?', 'Share your morning routine',
            'This or that: Classic vs Modern', 'Weekend plans check-in',
            'Rate our new packaging 1-10', 'Fill in the blank challenge',
            'Guess the upcoming product hint', 'Tell us your wishlist item',
        ],
    }

    random.seed(42)
    start_date = date(2025, 1, 1)
    end_date = date(2025, 3, 31)
    current_date = start_date
    row = 2

    while current_date <= end_date:
        # 2-4 posts per day across platforms
        num_posts = random.randint(2, 4)
        selected_platforms = random.sample(platforms, num_posts)

        for platform in selected_platforms:
            ct = random.choice(content_types)
            topic = random.choice(topics[ct])
            campaign = random.choice(campaigns)

            ws1.cell(row=row, column=1, value=current_date.strftime('%Y-%m-%d'))
            ws1.cell(row=row, column=2, value=platform)
            ws1.cell(row=row, column=3, value=ct)
            ws1.cell(row=row, column=4, value=topic)
            ws1.cell(row=row, column=5, value=campaign)
            row += 1

        current_date += timedelta(days=1)

    # Auto-size columns
    ws1.column_dimensions['A'].width = 14
    ws1.column_dimensions['B'].width = 14
    ws1.column_dimensions['C'].width = 22
    ws1.column_dimensions['D'].width = 45
    ws1.column_dimensions['E'].width = 18

    # --- Sheet 2: Brand Guidelines ---
    ws2 = wb.create_sheet('Brand Guidelines')

    ws2['A1'] = 'Content Type Distribution Guidelines'
    ws2['A1'].font = Font(bold=True, size=14)

    ws2['A3'] = 'Content Type'
    ws2['B3'] = 'Target %'
    ws2['C3'] = 'Color Code'
    ws2['A3'].font = Font(bold=True)
    ws2['B3'].font = Font(bold=True)
    ws2['C3'].font = Font(bold=True)

    guidelines = [
        ('Product Posts', '25%', 'Blue (#4472C4)'),
        ('Educational Content', '20%', 'Green (#70AD47)'),
        ('Behind the Scenes', '15%', 'Yellow (#FFC000)'),
        ('Promotions', '15%', 'Orange (#ED7D31)'),
        ('User Generated Content', '15%', 'Purple (#7030A0)'),
        ('Engagement Questions', '10%', 'Pink (#FF69B4)'),
    ]
    for i, (ct, pct, color) in enumerate(guidelines, 4):
        ws2.cell(row=i, column=1, value=ct)
        ws2.cell(row=i, column=2, value=pct)
        ws2.cell(row=i, column=3, value=color)

    ws2['A11'] = 'Active Campaigns'
    ws2['A11'].font = Font(bold=True, size=12)
    ws2['A12'] = 'Campaign Name'
    ws2['B12'] = 'Start Date'
    ws2['C12'] = 'End Date'
    ws2['A12'].font = Font(bold=True)
    ws2['B12'].font = Font(bold=True)
    ws2['C12'].font = Font(bold=True)

    campaigns_data = [
        ('Spring Launch', '2025-02-15', '2025-03-15'),
        ('Earth Day', '2025-03-20', '2025-04-22'),
        ('Q1 Push', '2025-01-01', '2025-03-31'),
        ('Brand Awareness', '2025-01-15', '2025-02-28'),
    ]
    for i, (name, start, end) in enumerate(campaigns_data, 13):
        ws2.cell(row=i, column=1, value=name)
        ws2.cell(row=i, column=2, value=start)
        ws2.cell(row=i, column=3, value=end)

    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 20

    # --- Sheet 3: Platforms ---
    ws3 = wb.create_sheet('Platform Info')
    ws3['A1'] = 'Platform Posting Guidelines'
    ws3['A1'].font = Font(bold=True, size=14)

    ws3['A3'] = 'Platform'
    ws3['B3'] = 'Best Post Time'
    ws3['C3'] = 'Max Posts/Day'
    ws3['D3'] = 'Primary Audience'
    for col in range(1, 5):
        ws3.cell(row=3, column=col).font = Font(bold=True)

    platform_info = [
        ('Instagram', '11:00 AM - 1:00 PM', 2, 'Ages 18-34, visual-focused'),
        ('Facebook', '1:00 PM - 4:00 PM', 3, 'Ages 25-54, broad reach'),
        ('Twitter/X', '9:00 AM - 12:00 PM', 5, 'Ages 18-49, news-oriented'),
        ('LinkedIn', '7:30 AM - 8:30 AM', 1, 'Professionals 25-55'),
        ('TikTok', '7:00 PM - 9:00 PM', 3, 'Ages 16-34, entertainment'),
    ]
    for i, (plat, time_str, max_posts, audience) in enumerate(platform_info, 4):
        ws3.cell(row=i, column=1, value=plat)
        ws3.cell(row=i, column=2, value=time_str)
        ws3.cell(row=i, column=3, value=max_posts)
        ws3.cell(row=i, column=4, value=audience)

    ws3.column_dimensions['A'].width = 14
    ws3.column_dimensions['B'].width = 22
    ws3.column_dimensions['C'].width = 16
    ws3.column_dimensions['D'].width = 35

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
