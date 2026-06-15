"""
Reward Script: Multi-app workflow — PDF extraction to CSV, LibreOffice Calc analysis, PDF export
Task ID: pdf_cross_139
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: CSV file exists with 50 data rows and 4 correct columns           — 0.25 pts
  Component 2: XLSX Data sheet has 50 rows with correct headers                  — 0.20 pts
  Component 3: XLSX Statistics sheet has category stats (COUNT, SUM, AVERAGE)    — 0.20 pts
  Component 4: XLSX has a chart (in Statistics sheet)                            — 0.20 pts
  Component 5: data_analysis.pdf exists and is a valid PDF with content          — 0.15 pts
  Total: 1.0
"""

import os
import csv

WORKDIR = '/home/user'
TASK_ID = 'pdf_cross_139'

CSV_PATH   = '/home/user/Documents/data.csv'
XLSX_PATH  = '/home/user/pdf_cross_139.xlsx'
PDF_PATH   = '/home/user/Documents/data_analysis.pdf'

EXPECTED_CSV_HEADERS  = ['Date', 'Category', 'Amount', 'Status']
EXPECTED_CATEGORIES   = {'Technology', 'Healthcare', 'Finance', 'Retail', 'Education'}
EXPECTED_STAT_HEADERS = ['Category', 'COUNT', 'SUM', 'AVERAGE']


def verify_task():
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # -----------------------------------------------------------------------
    # Component 1: CSV file exists with 50 data rows and 4 correct columns
    # (0.25 points)
    # This FAILS on initial_env (file absent) and PASSES on golden_env.
    # -----------------------------------------------------------------------
    try:
        if not os.path.exists(CSV_PATH):
            print(f"FAIL: Component 1 — CSV file not found at {CSV_PATH}")
        else:
            with open(CSV_PATH, 'r', newline='') as f:
                reader = csv.reader(f)
                rows = list(reader)

            if len(rows) == 0:
                print("FAIL: Component 1 — CSV file is empty")
            else:
                header_row = [h.strip() for h in rows[0]]
                data_rows  = rows[1:]

                headers_ok  = (header_row == EXPECTED_CSV_HEADERS)
                row_count_ok = (len(data_rows) == 50)
                col_count_ok = all(len(r) == 4 for r in data_rows) if data_rows else False

                if headers_ok and row_count_ok and col_count_ok:
                    print(f"PASS: Component 1 — CSV has 50 data rows, headers {header_row} (0.25 pts)")
                    total_score += 0.25
                else:
                    issues = []
                    if not headers_ok:
                        issues.append(f"headers {header_row} != {EXPECTED_CSV_HEADERS}")
                    if not row_count_ok:
                        issues.append(f"{len(data_rows)} data rows (expected 50)")
                    if not col_count_ok:
                        issues.append("some rows do not have 4 columns")
                    print(f"FAIL: Component 1 — CSV issues: {'; '.join(issues)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: XLSX Data sheet has 50 data rows with correct headers
    # (0.20 points)
    # This FAILS on initial_env (file absent) and PASSES on golden_env.
    # -----------------------------------------------------------------------
    wb = None
    try:
        if not os.path.exists(XLSX_PATH):
            print(f"FAIL: Component 2 — XLSX file not found at {XLSX_PATH}")
        else:
            import openpyxl
            wb = openpyxl.load_workbook(XLSX_PATH)

            if 'Data' not in wb.sheetnames:
                print(f"FAIL: Component 2 — No 'Data' sheet found. Sheets: {wb.sheetnames}")
            else:
                ws_data = wb['Data']
                # Read header row
                headers = [ws_data.cell(row=1, column=c).value for c in range(1, 5)]
                data_row_count = ws_data.max_row - 1  # subtract header row

                headers_ok    = (headers == EXPECTED_CSV_HEADERS)
                row_count_ok  = (data_row_count == 50)

                if headers_ok and row_count_ok:
                    print(f"PASS: Component 2 — Data sheet has headers {headers} and {data_row_count} data rows (0.20 pts)")
                    total_score += 0.20
                else:
                    issues = []
                    if not headers_ok:
                        issues.append(f"headers {headers} != {EXPECTED_CSV_HEADERS}")
                    if not row_count_ok:
                        issues.append(f"{data_row_count} data rows (expected 50)")
                    print(f"FAIL: Component 2 — Data sheet issues: {'; '.join(issues)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: XLSX Statistics sheet has summary stats by category
    # (COUNT, SUM, AVERAGE for all 5 categories + TOTAL row)
    # (0.20 points)
    # This FAILS on initial_env (file absent) and PASSES on golden_env.
    # -----------------------------------------------------------------------
    try:
        if wb is None and os.path.exists(XLSX_PATH):
            import openpyxl
            wb = openpyxl.load_workbook(XLSX_PATH)

        if wb is None:
            print(f"FAIL: Component 3 — XLSX file not accessible")
        elif 'Statistics' not in wb.sheetnames:
            print(f"FAIL: Component 3 — No 'Statistics' sheet found. Sheets: {wb.sheetnames}")
        else:
            ws_stats = wb['Statistics']

            # Find the header row containing Category/COUNT/SUM/AVERAGE
            header_row_num = None
            for r in range(1, ws_stats.max_row + 1):
                row_vals = [ws_stats.cell(row=r, column=c).value for c in range(1, 5)]
                if row_vals[0] == 'Category':
                    header_row_num = r
                    break

            if header_row_num is None:
                print("FAIL: Component 3 — Statistics sheet has no 'Category' header row")
            else:
                # Collect category rows after the header
                categories_found = set()
                total_row_count  = 0

                # Validate column headers using boolean expressions (no bare '= True')
                stat_headers = [ws_stats.cell(row=header_row_num, column=c).value for c in range(1, 5)]
                has_count_col = (stat_headers[1] is not None and
                                 str(stat_headers[1]).upper() in ('COUNT', 'CNT'))
                has_sum_col   = (stat_headers[2] is not None and
                                 str(stat_headers[2]).upper() in ('SUM', 'TOTAL'))
                has_avg_col   = (stat_headers[3] is not None and
                                 str(stat_headers[3]).upper() in ('AVERAGE', 'AVG', 'MEAN'))

                for r in range(header_row_num + 1, ws_stats.max_row + 1):
                    cat_val = ws_stats.cell(row=r, column=1).value
                    if cat_val is None:
                        continue
                    cat_str = str(cat_val).strip()
                    if cat_str.upper() == 'TOTAL':
                        total_row_count += 1
                    elif cat_str in EXPECTED_CATEGORIES:
                        # Verify numeric values for COUNT, SUM, AVG columns
                        count_val = ws_stats.cell(row=r, column=2).value
                        sum_val   = ws_stats.cell(row=r, column=3).value
                        avg_val   = ws_stats.cell(row=r, column=4).value
                        if (count_val is not None and sum_val is not None and avg_val is not None):
                            categories_found.add(cat_str)

                all_cats_present = (categories_found == EXPECTED_CATEGORIES)
                cols_ok          = (has_count_col and has_sum_col and has_avg_col)
                has_total_row    = (total_row_count >= 1)

                if all_cats_present and has_total_row and cols_ok:
                    print(f"PASS: Component 3 — Statistics sheet has COUNT/SUM/AVERAGE for all 5 categories + TOTAL row (0.20 pts)")
                    total_score += 0.20
                else:
                    issues = []
                    if not all_cats_present:
                        missing = EXPECTED_CATEGORIES - categories_found
                        issues.append(f"missing categories: {missing}")
                    if not has_total_row:
                        issues.append(f"no TOTAL row (total_row_count={total_row_count})")
                    if not cols_ok:
                        issues.append(f"stat column headers issue: {stat_headers}")
                    print(f"FAIL: Component 3 — Statistics sheet issues: {'; '.join(issues)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: XLSX has a chart (in Statistics sheet or any sheet)
    # (0.20 points)
    # This FAILS on initial_env (file absent) and PASSES on golden_env.
    # -----------------------------------------------------------------------
    try:
        if wb is None and os.path.exists(XLSX_PATH):
            import openpyxl
            wb = openpyxl.load_workbook(XLSX_PATH)

        if wb is None:
            print(f"FAIL: Component 4 — XLSX file not accessible")
        else:
            total_charts = 0
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                charts = ws._charts
                total_charts += len(charts)

            if total_charts >= 1:
                print(f"PASS: Component 4 — Found {total_charts} chart(s) in workbook (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 4 — No charts found in workbook (expected at least 1)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -----------------------------------------------------------------------
    # Component 5: data_analysis.pdf exists and is a valid PDF with content
    # (0.15 points)
    # This FAILS on initial_env (file absent) and PASSES on golden_env.
    # -----------------------------------------------------------------------
    try:
        if not os.path.exists(PDF_PATH):
            print(f"FAIL: Component 5 — PDF file not found at {PDF_PATH}")
        else:
            pdf_size = os.path.getsize(PDF_PATH)
            if pdf_size < 100:
                print(f"FAIL: Component 5 — PDF file is too small ({pdf_size} bytes), likely empty/corrupt")
            else:
                # Verify it starts with the PDF magic bytes
                with open(PDF_PATH, 'rb') as f:
                    header = f.read(5)
                if header == b'%PDF-':
                    print(f"PASS: Component 5 — data_analysis.pdf exists, is a valid PDF, size={pdf_size} bytes (0.15 pts)")
                    total_score += 0.15
                else:
                    print(f"FAIL: Component 5 — File at {PDF_PATH} does not start with PDF magic bytes (got {header!r})")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # -----------------------------------------------------------------------
    # Final score
    # -----------------------------------------------------------------------
    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


if __name__ == '__main__':
    verify_task()
