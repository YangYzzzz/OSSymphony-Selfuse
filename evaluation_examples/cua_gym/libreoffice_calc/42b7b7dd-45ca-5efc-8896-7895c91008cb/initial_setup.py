"""
Initial Setup: Set page header/footer margins in LibreOffice Calc
Task ID: calc_gfl_078
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.worksheet.header_footer import HeaderFooterItem

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_078'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Report ---
    ws = wb.active
    ws.title = 'Report'

    # Headers
    headers = ['Date', 'Description', 'Category', 'Debit ($)', 'Credit ($)', 'Balance ($)']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal='center', vertical='center')
    white_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Financial data - 35 rows of realistic transaction data
    data = [
        ['2025-01-02', 'Opening Balance', 'Balance', None, None, 125000.00],
        ['2025-01-03', 'Office Rent Payment - Q1', 'Rent', 4500.00, None, 120500.00],
        ['2025-01-05', 'Client Invoice #1042 - Meridian Corp', 'Revenue', None, 15200.00, 135700.00],
        ['2025-01-08', 'Software Licenses - Annual Renewal', 'Software', 2340.00, None, 133360.00],
        ['2025-01-10', 'Employee Payroll - January Cycle 1', 'Payroll', 28750.00, None, 104610.00],
        ['2025-01-12', 'Client Invoice #1043 - Apex Solutions', 'Revenue', None, 8900.00, 113510.00],
        ['2025-01-14', 'Utility Bills - Electricity & Water', 'Utilities', 876.50, None, 112633.50],
        ['2025-01-15', 'Marketing Campaign - Social Media Ads', 'Marketing', 3200.00, None, 109433.50],
        ['2025-01-17', 'Client Invoice #1044 - TechVista Ltd', 'Revenue', None, 22500.00, 131933.50],
        ['2025-01-19', 'Office Supplies - Stationery & Equipment', 'Supplies', 1245.75, None, 130687.75],
        ['2025-01-20', 'Insurance Premium - Q1', 'Insurance', 3600.00, None, 127087.75],
        ['2025-01-22', 'Client Invoice #1045 - Greenfield Inc', 'Revenue', None, 11350.00, 138437.75],
        ['2025-01-24', 'Employee Payroll - January Cycle 2', 'Payroll', 28750.00, None, 109687.75],
        ['2025-01-25', 'Travel Expenses - Sales Team Conference', 'Travel', 4820.00, None, 104867.75],
        ['2025-01-27', 'Client Invoice #1046 - Summit Partners', 'Revenue', None, 19750.00, 124617.75],
        ['2025-01-28', 'IT Infrastructure Maintenance', 'IT', 2100.00, None, 122517.75],
        ['2025-01-30', 'Professional Services - Legal Counsel', 'Legal', 5500.00, None, 117017.75],
        ['2025-02-01', 'Client Invoice #1047 - BlueStar Dynamics', 'Revenue', None, 13400.00, 130417.75],
        ['2025-02-03', 'Office Rent Payment - February', 'Rent', 4500.00, None, 125917.75],
        ['2025-02-05', 'Telecommunications - Phone & Internet', 'Utilities', 542.00, None, 125375.75],
        ['2025-02-07', 'Client Invoice #1048 - Cascade Systems', 'Revenue', None, 16800.00, 142175.75],
        ['2025-02-10', 'Employee Payroll - February Cycle 1', 'Payroll', 29100.00, None, 113075.75],
        ['2025-02-12', 'Training & Development Workshop', 'Training', 2750.00, None, 110325.75],
        ['2025-02-14', 'Client Invoice #1049 - Pinnacle Group', 'Revenue', None, 9200.00, 119525.75],
        ['2025-02-17', 'Equipment Lease - Copier & Printers', 'Equipment', 1800.00, None, 117725.75],
        ['2025-02-19', 'Client Invoice #1050 - Nova Enterprises', 'Revenue', None, 27300.00, 145025.75],
        ['2025-02-21', 'Cleaning & Maintenance Services', 'Maintenance', 950.00, None, 144075.75],
        ['2025-02-24', 'Employee Payroll - February Cycle 2', 'Payroll', 29100.00, None, 114975.75],
        ['2025-02-25', 'Client Invoice #1051 - Horizon Analytics', 'Revenue', None, 14600.00, 129575.75],
        ['2025-02-26', 'Advertising - Print Media', 'Marketing', 1850.00, None, 127725.75],
        ['2025-02-27', 'Accounting & Audit Fees', 'Professional', 4200.00, None, 123525.75],
        ['2025-02-28', 'Client Invoice #1052 - Vertex Solutions', 'Revenue', None, 18500.00, 142025.75],
        ['2025-03-01', 'Office Rent Payment - March', 'Rent', 4500.00, None, 137525.75],
        ['2025-03-03', 'Client Invoice #1053 - Crestwood Ltd', 'Revenue', None, 21000.00, 158525.75],
        ['2025-03-05', 'Miscellaneous - Courier & Postage', 'Miscellaneous', 385.25, None, 158140.50],
    ]

    currency_fmt = '$#,##0.00'
    date_fmt = 'yyyy-mm-dd'
    thin = Side(style='thin', color='D9D9D9')
    border = Border(bottom=thin)

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            if c == 1:
                cell.number_format = date_fmt
            elif c in (4, 5, 6):
                cell.number_format = currency_fmt

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 16

    # Configure header and footer
    ws.oddHeader.left.text = "Northwind Financial Services"
    ws.oddHeader.left.font = "Calibri,Bold"
    ws.oddHeader.left.size = 12
    ws.oddHeader.right.text = "&D"  # Current date
    ws.oddHeader.right.font = "Calibri,Regular"
    ws.oddHeader.right.size = 10

    ws.oddFooter.center.text = "Page &P of &N"
    ws.oddFooter.center.font = "Calibri,Regular"
    ws.oddFooter.center.size = 10

    # Page setup - use default header/footer margins (NOT the target values)
    # Default header margin is typically ~0.3 inches (~0.76cm), footer ~0.3 inches
    # We leave them at defaults so the task is meaningful
    ws.page_margins.header = 0.3  # ~0.76cm - default, NOT the target 1.5cm
    ws.page_margins.footer = 0.3  # ~0.76cm - default, NOT the target 1.0cm

    # Standard page margins
    ws.page_margins.left = 0.7
    ws.page_margins.right = 0.7
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75

    # Print settings
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_area = 'A1:F36'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
