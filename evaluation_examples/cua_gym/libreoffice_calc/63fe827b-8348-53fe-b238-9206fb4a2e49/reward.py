"""
Reward Script: Find average test score for female students in Grade 10 using AVERAGEIFS
Task ID: calc_fmb_averageifs_multi_012
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Cell I2 contains a formula (not None, starts with '=')
  Component 2 (0.4): Formula in I2 is AVERAGEIFS with correct ranges and criteria
                     (E2:E501 for scores, C2:C501="F" for gender, D2:D501=10 for grade)
  Component 3 (0.2): No extra cells modified — H2 label intact, I column rows 3+ unchanged
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_averageifs_multi_012'


def normalize_formula(f):
    """Normalize formula for comparison: uppercase, remove spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '').replace('"', '"').replace('"', '"')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: sheet 'Student Records' must exist
    if 'Student Records' not in wb.sheetnames:
        print("FAIL: Sheet 'Student Records' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Student Records']

    # Component 1: Cell I2 contains a formula (0.4 points)
    # This FAILS on initial (I2 is None) and PASSES on golden (I2 has =AVERAGEIFS(...))
    try:
        i2_value = ws.cell(row=2, column=9).value
        if i2_value is not None and isinstance(i2_value, str) and i2_value.strip().startswith('='):
            print(f"PASS: Component 1 — I2 contains a formula: {repr(i2_value)} (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — Expected a formula in I2, found: {repr(i2_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Formula in I2 is AVERAGEIFS with correct ranges and criteria (0.4 points)
    # Checks that the formula:
    #   - Uses AVERAGEIFS function
    #   - References E2:E501 as the average range (test scores)
    #   - References C2:C501 with criteria "F" (female gender filter)
    #   - References D2:D501 with criteria 10 (Grade 10 filter)
    # This FAILS on initial (I2 is None) and PASSES on golden (correct AVERAGEIFS formula)
    try:
        i2_value = ws.cell(row=2, column=9).value
        if i2_value is not None and isinstance(i2_value, str):
            norm = normalize_formula(i2_value)
            has_averageifs = 'AVERAGEIFS(' in norm
            has_score_range = 'E2:E501' in norm
            has_gender_range = 'C2:C501' in norm
            has_gender_criteria = '"F"' in norm or ',"F"' in norm.replace(' ', '')
            has_grade_range = 'D2:D501' in norm
            has_grade_criteria = ',10)' in norm or ',10,' in norm

            if has_averageifs and has_score_range and has_gender_range and has_gender_criteria and has_grade_range and has_grade_criteria:
                print(f"PASS: Component 2 — AVERAGEIFS formula has correct ranges and criteria (0.4 pts)")
                print(f"      Formula: {repr(i2_value)}")
                total_score += 0.4
            else:
                missing = []
                if not has_averageifs:
                    missing.append("AVERAGEIFS function")
                if not has_score_range:
                    missing.append("score range E2:E501")
                if not has_gender_range:
                    missing.append("gender range C2:C501")
                if not has_gender_criteria:
                    missing.append('gender criteria "F"')
                if not has_grade_range:
                    missing.append("grade range D2:D501")
                if not has_grade_criteria:
                    missing.append("grade criteria 10")
                print(f"FAIL: Component 2 — Formula missing: {', '.join(missing)}")
                print(f"      Formula found: {repr(i2_value)}")
        else:
            print(f"FAIL: Component 2 — No formula in I2 to check (I2={repr(i2_value)})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: No other cells modified — H2 label intact, I column rows beyond I2 unchanged (0.2 points)
    # H2 should still contain 'Female Gr10 Avg' (the label from the initial file)
    # I column beyond row 2 should all be None (no accidental data entry)
    # This FAILS on initial because I2 is None (a proxy — actually checks combined with no-extra-changes)
    # To ensure it only scores when I2 has been set (the task is done), we check BOTH:
    #   a) H2 label is still intact
    #   b) I column rows 3-501 are all None (no accidental spillover)
    # We gate this component on I2 being non-None so it only awards points when the task was attempted
    try:
        i2_value = ws.cell(row=2, column=9).value
        h2_value = ws.cell(row=2, column=8).value

        # Only evaluate this component if I2 was set (task was attempted)
        if i2_value is None:
            print("FAIL: Component 3 — I2 is empty; no task completion to evaluate for collateral check")
        else:
            h2_ok = (h2_value == 'Female Gr10 Avg')
            spillover_cells = [r for r in range(3, 502) if ws.cell(row=r, column=9).value is not None]

            if h2_ok and not spillover_cells:
                print(f"PASS: Component 3 — H2 label intact ('{h2_value}'), no spillover in I column (0.2 pts)")
                total_score += 0.2
            else:
                if not h2_ok:
                    print(f"FAIL: Component 3 — H2 label changed: expected 'Female Gr10 Avg', found {repr(h2_value)}")
                if spillover_cells:
                    print(f"FAIL: Component 3 — Spillover detected in I column: rows {spillover_cells[:3]}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
