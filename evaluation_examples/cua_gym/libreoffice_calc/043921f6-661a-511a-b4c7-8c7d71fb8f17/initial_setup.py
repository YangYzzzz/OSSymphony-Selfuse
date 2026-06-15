"""
Initial Setup: Charts sheet with two embedded charts and data tables
Task ID: calc_mcp_085
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_085'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Charts (the main sheet with data + charts) ---
    ws = wb.active
    ws.title = 'Charts'

    # Style definitions
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='B0B0B0'),
        right=Side(style='thin', color='B0B0B0'),
        top=Side(style='thin', color='B0B0B0'),
        bottom=Side(style='thin', color='B0B0B0'),
    )

    # Data Table 1: Monthly Revenue
    headers1 = ['Month', 'Revenue ($)', 'Expenses ($)', 'Profit ($)']
    for c, h in enumerate(headers1, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    monthly_data = [
        ['January',   45230, 31200, 14030],
        ['February',  52180, 33400, 18780],
        ['March',     48900, 29800, 19100],
        ['April',     61340, 35600, 25740],
        ['May',       55720, 32100, 23620],
        ['June',      67890, 38400, 29490],
        ['July',      72100, 41200, 30900],
        ['August',    68450, 39500, 28950],
        ['September', 59300, 34800, 24500],
        ['October',   63200, 36700, 26500],
        ['November',  71500, 40100, 31400],
        ['December',  78900, 43200, 35700],
    ]
    for r, row_data in enumerate(monthly_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c >= 2:
                cell.number_format = '#,##0'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 13

    # Data Table 2: Department Summary (below, rows 16+)
    dept_headers = ['Department', 'Headcount', 'Budget ($)', 'Utilization (%)']
    for c, h in enumerate(dept_headers, 1):
        cell = ws.cell(row=16, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    dept_data = [
        ['Engineering',  42, 2850000, 87.3],
        ['Marketing',    18, 1200000, 92.1],
        ['Sales',        25, 1650000, 78.5],
        ['Operations',   15,  980000, 83.7],
        ['HR',           8,   520000, 71.2],
        ['Finance',      12,  780000, 88.9],
        ['Legal',        6,   450000, 65.4],
        ['Support',      20, 1100000, 90.6],
    ]
    for r, row_data in enumerate(dept_data, 17):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 3:
                cell.number_format = '#,##0'
            elif c == 4:
                cell.number_format = '0.0'

    # --- Chart 1: Bar chart for Monthly Revenue/Expenses (anchored at F2) ---
    bar_chart = BarChart()
    bar_chart.type = 'col'
    bar_chart.title = 'Monthly Revenue vs Expenses'
    bar_chart.y_axis.title = 'Amount ($)'
    bar_chart.x_axis.title = 'Month'
    bar_chart.width = 18
    bar_chart.height = 12
    data_ref = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=13)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=13)
    bar_chart.add_data(data_ref, titles_from_data=True)
    bar_chart.set_categories(cats_ref)
    ws.add_chart(bar_chart, 'F2')

    # --- Chart 2: Line chart for Profit Trend (anchored at F18) ---
    line_chart = LineChart()
    line_chart.title = 'Monthly Profit Trend'
    line_chart.y_axis.title = 'Profit ($)'
    line_chart.x_axis.title = 'Month'
    line_chart.width = 18
    line_chart.height = 12
    profit_ref = Reference(ws, min_col=4, min_row=1, max_row=13)
    cats_ref2 = Reference(ws, min_col=1, min_row=2, max_row=13)
    line_chart.add_data(profit_ref, titles_from_data=True)
    line_chart.set_categories(cats_ref2)
    ws.add_chart(line_chart, 'F18')

    # --- Sheet 2: Raw Data (supporting data sheet) ---
    ws2 = wb.create_sheet('Raw Data')
    raw_headers = ['Date', 'Transaction ID', 'Category', 'Amount', 'Status']
    for c, h in enumerate(raw_headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)

    import random
    random.seed(42)
    categories = ['Product Sales', 'Service Revenue', 'Licensing', 'Consulting']
    statuses = ['Completed', 'Pending', 'Completed', 'Completed', 'Refunded']
    for r in range(2, 22):
        ws2.cell(row=r, column=1, value=f'2025-{(r % 12) + 1:02d}-{(r * 3 % 28) + 1:02d}')
        ws2.cell(row=r, column=2, value=f'TXN-{10000 + r}')
        ws2.cell(row=r, column=3, value=categories[r % len(categories)])
        ws2.cell(row=r, column=4, value=round(random.uniform(500, 15000), 2))
        ws2.cell(row=r, column=5, value=statuses[r % len(statuses)])

    # NO special print settings — default state
    # This is what the agent needs to configure

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
