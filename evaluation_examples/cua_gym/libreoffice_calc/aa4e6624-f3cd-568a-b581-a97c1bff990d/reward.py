"""
Reward Script: Customer Order Tracking Sheet with Status Workflow
Task ID: calc_gpm_076
Domain: libreoffice_calc
Scoring:
  Component 1: Days in Process formulas (I4:I15) — 0.20
  Component 2: Order Summary section (A17:D17 merge, labels, COUNTIF formulas) — 0.25
  Component 3: Revenue calculation (E17 label + E18 SUMIF formula) — 0.15
  Component 4: Doughnut chart present with correct title — 0.15
  Component 5: Conditional formatting on H column (status colors) — 0.15
  Component 6: Conditional formatting on I column (processing time alerts) — 0.10
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_076'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Orders' sheet must exist
    if 'Orders' not in wb.sheetnames:
        print("FAIL: 'Orders' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Orders']

    # Component 1: Days in Process formulas in I4:I15 (0.20 points)
    # Initial has None in I4:I15; golden has =TODAY()-F formulas
    try:
        formula_count = 0
        for row in range(4, 16):
            val = ws.cell(row=row, column=9).value  # column I
            if val is not None and isinstance(val, str) and '=TODAY()' in val.upper() and 'F' in val.upper():
                formula_count += 1
        if formula_count >= 10:
            print(f"PASS: Component 1 -- Days in Process formulas found in {formula_count}/12 cells (0.20 pts)")
            total_score += 0.20
        elif formula_count >= 6:
            partial = 0.10
            print(f"PARTIAL: Component 1 -- Only {formula_count}/12 formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 -- Only {formula_count}/12 Days in Process formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2a: Order Summary merge + label (0.08 points)
    # Golden has A17:D17 merged with "Order Summary" bold
    try:
        a17_merge_found = any('A17' in str(mr) and 'D17' in str(mr) for mr in ws.merged_cells.ranges)
        a17_val = ws['A17'].value
        if a17_merge_found and a17_val and 'Order Summary' in str(a17_val):
            print(f"PASS: Component 2a -- A17:D17 merged with 'Order Summary' (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 2a -- A17 merge/value: merged={a17_merge_found}, value={repr(a17_val)}")
    except Exception as e:
        print(f"ERROR: Component 2a -- {e}")

    # Component 2b: A17 bold (0.02 points)
    try:
        if ws['A17'].font.bold:
            print(f"PASS: Component 2b -- A17 is bold (0.02 pts)")
            total_score += 0.02
        else:
            print(f"FAIL: Component 2b -- A17 not bold")
    except Exception as e:
        print(f"ERROR: Component 2b -- {e}")

    # Component 2c: COUNTIF formulas in rows 18-23 (0.15 points)
    try:
        expected_statuses = ['Received', 'Processing', 'Shipped', 'Delivered', 'Returned', 'Cancelled']
        countif_found = 0
        for row in range(18, 24):
            label = ws.cell(row=row, column=1).value
            formula = ws.cell(row=row, column=2).value
            if label and formula and isinstance(formula, str) and 'COUNTIF' in formula.upper():
                if any(s.lower() in str(label).lower() for s in expected_statuses):
                    countif_found += 1

        if countif_found >= 5:
            print(f"PASS: Component 2c -- {countif_found}/6 COUNTIF status formulas found (0.15 pts)")
            total_score += 0.15
        elif countif_found >= 3:
            print(f"PARTIAL: Component 2c -- {countif_found}/6 COUNTIF status formulas found (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 2c -- Only {countif_found}/6 COUNTIF formulas found")
    except Exception as e:
        print(f"ERROR: Component 2c -- {e}")

    # Component 3a: Revenue label (0.05 points)
    try:
        e17_val = ws['E17'].value
        if e17_val and 'Total Revenue' in str(e17_val):
            print(f"PASS: Component 3a -- E17 contains 'Total Revenue' label (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 3a -- E17={repr(e17_val)}, expected 'Total Revenue:'")
    except Exception as e:
        print(f"ERROR: Component 3a -- {e}")

    # Component 3b: SUMIF revenue formula (0.10 points)
    try:
        e18_val = ws['E18'].value
        if e18_val and isinstance(e18_val, str) and 'SUMIF' in e18_val.upper():
            print(f"PASS: Component 3b -- E18 contains SUMIF formula: {e18_val} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3b -- E18={repr(e18_val)}, expected SUMIF formula")
    except Exception as e:
        print(f"ERROR: Component 3b -- {e}")

    # Component 4: Doughnut chart with title (0.15 points)
    # Initial has no charts; golden has a doughnut chart titled 'Order Status Breakdown'
    try:
        charts = ws._charts
        if len(charts) >= 1:
            # Check for doughnut chart type and title
            doughnut_count = sum(1 for ch in charts if 'Doughnut' in type(ch).__name__)
            title_match_count = 0
            for ch in charts:
                if 'Doughnut' in type(ch).__name__:
                    try:
                        paras = ch.title.tx.rich.paragraphs
                        for para in paras:
                            if hasattr(para, 'r') and para.r:
                                for run in para.r:
                                    if run.t and 'Order Status' in run.t:
                                        title_match_count += 1
                    except Exception:
                        pass

            if doughnut_count >= 1 and title_match_count >= 1:
                print(f"PASS: Component 4 -- Doughnut chart with correct title found (0.15 pts)")
                total_score += 0.15
            elif doughnut_count >= 1:
                print(f"PARTIAL: Component 4 -- Doughnut chart found but title mismatch (0.08 pts)")
                total_score += 0.08
            elif len(charts) >= 1:
                # At least a chart exists even if not doughnut type
                print(f"PARTIAL: Component 4 -- Chart found but not doughnut type: {type(charts[0]).__name__} (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 4 -- No charts found")
        else:
            print(f"FAIL: Component 4 -- No charts found")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # Component 5: Conditional formatting on H column (status colors) (0.15 points)
    # Initial has no conditional formatting; golden has 6 rules on H4:H15
    try:
        h_cf_rules = 0
        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            if 'H' in cf_range:
                for rule in cf.rules:
                    h_cf_rules += 1

        if h_cf_rules >= 5:
            print(f"PASS: Component 5 -- {h_cf_rules} conditional formatting rules on H column (0.15 pts)")
            total_score += 0.15
        elif h_cf_rules >= 3:
            partial = 0.08
            print(f"PARTIAL: Component 5 -- {h_cf_rules} CF rules on H column ({partial} pts)")
            total_score += partial
        elif h_cf_rules >= 1:
            partial = 0.05
            print(f"PARTIAL: Component 5 -- {h_cf_rules} CF rules on H column ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 -- No conditional formatting on H column")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    # Component 6: Conditional formatting on I column (processing time alerts) (0.10 points)
    # Initial has no CF on I; golden has 3 rules on I4:I15 (>7 red, 4-7 orange, <=3 green)
    try:
        i_cf_rules = 0
        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            if 'I' in cf_range and 'H' not in cf_range:
                for rule in cf.rules:
                    i_cf_rules += 1

        if i_cf_rules >= 3:
            print(f"PASS: Component 6 -- {i_cf_rules} conditional formatting rules on I column (0.10 pts)")
            total_score += 0.10
        elif i_cf_rules >= 1:
            partial = 0.05
            print(f"PARTIAL: Component 6 -- {i_cf_rules} CF rules on I column ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 6 -- No conditional formatting on I column")
    except Exception as e:
        print(f"ERROR: Component 6 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
