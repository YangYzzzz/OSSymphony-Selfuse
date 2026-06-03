"""
Initial Setup: Customer order tracking sheet with status workflow and delivery timeline formatting.
Task ID: calc_gpm_076
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_076'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"

    # --- Title row: merge A1:I1 ---
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "Customer Order Tracking - April 2026"
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="FFCC6600", end_color="FFCC6600", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Row 3 headers ---
    headers = ["Order #", "Customer", "Product", "Qty", "Total",
               "Order Date", "Ship Date", "Status", "Days in Process"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FFCC6600", end_color="FFCC6600", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Order data rows 4-15 ---
    orders = [
        ["ORD-1001", "Sarah Chen", "Laptop Pro 15\"", 2, 2499.98,
         date(2026, 3, 18), date(2026, 3, 21), "Delivered"],
        ["ORD-1002", "Marcus Johnson", "Wireless Mouse", 5, 149.95,
         date(2026, 3, 20), date(2026, 3, 23), "Delivered"],
        ["ORD-1003", "Emily Rodriguez", "Monitor 27\" 4K", 1, 549.99,
         date(2026, 3, 22), date(2026, 3, 26), "Shipped"],
        ["ORD-1004", "James Liu", "Keyboard Mechanical", 3, 389.97,
         date(2026, 3, 24), None, "Processing"],
        ["ORD-1005", "Olivia Patel", "USB-C Hub", 10, 399.90,
         date(2026, 3, 25), date(2026, 3, 28), "Shipped"],
        ["ORD-1006", "David Kim", "Webcam HD", 4, 319.96,
         date(2026, 3, 26), None, "Received"],
        ["ORD-1007", "Rachel Thompson", "Headset Wireless", 2, 179.98,
         date(2026, 3, 27), date(2026, 3, 30), "Delivered"],
        ["ORD-1008", "Ahmed Hassan", "Laptop Stand", 6, 239.94,
         date(2026, 3, 28), None, "Processing"],
        ["ORD-1009", "Lisa Wang", "External SSD 1TB", 3, 269.97,
         date(2026, 3, 29), date(2026, 4, 1), "Shipped"],
        ["ORD-1010", "Carlos Mendez", "Desk Mat XL", 8, 199.92,
         date(2026, 3, 30), None, "Cancelled"],
        ["ORD-1011", "Nicole Brown", "Charging Dock", 2, 139.98,
         date(2026, 3, 31), None, "Returned"],
        ["ORD-1012", "Tyler Jackson", "Portable Speaker", 1, 89.99,
         date(2026, 4, 1), None, "Received"],
    ]

    for r, order in enumerate(orders, 4):
        for c, val in enumerate(order, 1):
            cell = ws.cell(row=r, column=c, value=val)

    # --- Format columns ---
    # D (Qty) centered
    for row in range(4, 16):
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")

    # E (Total) $#,##0.00 right-aligned
    for row in range(4, 16):
        cell = ws.cell(row=row, column=5)
        cell.number_format = '$#,##0.00'
        cell.alignment = Alignment(horizontal="right")

    # F:G date format MMM DD
    for row in range(4, 16):
        for col in [6, 7]:
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.number_format = 'MMM DD'

    # --- Data validation on H4:H15 ---
    dv = DataValidation(
        type="list",
        formula1='"Received,Processing,Shipped,Delivered,Returned,Cancelled"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "Please select a valid status"
    dv.errorTitle = "Invalid Status"
    dv.prompt = "Select order status"
    dv.promptTitle = "Status"
    dv.add("H4:H15")
    ws.add_data_validation(dv)

    # --- Column widths ---
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 16

    # Row 1 height
    ws.row_dimensions[1].height = 30

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
