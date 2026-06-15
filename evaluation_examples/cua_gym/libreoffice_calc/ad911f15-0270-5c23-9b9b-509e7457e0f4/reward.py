"""
Reward Script: Add cell comments to HR onboarding checklist
Task ID: calc_hr_onboarding_checklist_comments_029
Domain: libreoffice_calc
Scoring:
  Component 1: B3 has I-9 identity documents comment (0.25 pts)
  Component 2: B7 has payroll/ADP portal comment (0.25 pts)
  Component 3: B11 has benefits enrollment comment (0.25 pts)
  Component 4: B15 has IT equipment setup comment (0.25 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_hr_onboarding_checklist_comments_029'
SHEET_NAME = 'Onboarding Checklist'

# Expected comment texts from task context (exact match required)
EXPECTED_COMMENTS = {
    'B3':  'Employee must provide original identity documents. See I-9 Policy Doc (HR-POL-001).',
    'B7':  'Submit bank details via ADP portal within 3 business days of start date.',
    'B11': 'Benefits enrollment window is 30 days from start date. Contact benefits@company.com for assistance.',
    'B15': 'Submit IT request form at least 5 days before start date. Laptop provisioning takes 3-4 business days.',
}


def normalize_comment(text):
    """Normalize comment text for comparison — strip whitespace and normalize newlines."""
    if text is None:
        return ''
    return text.strip().replace('\r\n', '\n').replace('\r', '\n').strip()


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Checks that cell comments were added to B3, B7, B11, and B15
    with the exact policy notes as specified in the task.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition gate: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: required sheet must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Component 1: B3 must have the I-9 identity documents comment (0.25 points)
    # This FAILS on initial (no comment) and PASSES on golden (comment present)
    try:
        cell_b3 = ws['B3']
        expected_b3 = EXPECTED_COMMENTS['B3']
        if cell_b3.comment is not None:
            actual_text = normalize_comment(cell_b3.comment.text)
            expected_text = normalize_comment(expected_b3)
            if expected_text in actual_text or actual_text == expected_text:
                print(f"PASS: Component 1 — B3 has I-9 comment: {repr(actual_text[:60])}... (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 1 — B3 comment text mismatch.")
                print(f"  Expected: {repr(expected_b3[:80])}")
                print(f"  Found:    {repr(actual_text[:80])}")
        else:
            print(f"FAIL: Component 1 — B3 has no comment (expected I-9 policy note)")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not check B3 comment: {e}")

    # Component 2: B7 must have the payroll/ADP portal comment (0.25 points)
    # This FAILS on initial (no comment) and PASSES on golden (comment present)
    try:
        cell_b7 = ws['B7']
        expected_b7 = EXPECTED_COMMENTS['B7']
        if cell_b7.comment is not None:
            actual_text = normalize_comment(cell_b7.comment.text)
            expected_text = normalize_comment(expected_b7)
            if expected_text in actual_text or actual_text == expected_text:
                print(f"PASS: Component 2 — B7 has payroll/ADP comment: {repr(actual_text[:60])}... (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 2 — B7 comment text mismatch.")
                print(f"  Expected: {repr(expected_b7[:80])}")
                print(f"  Found:    {repr(actual_text[:80])}")
        else:
            print(f"FAIL: Component 2 — B7 has no comment (expected payroll/ADP note)")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not check B7 comment: {e}")

    # Component 3: B11 must have the benefits enrollment comment (0.25 points)
    # This FAILS on initial (no comment) and PASSES on golden (comment present)
    try:
        cell_b11 = ws['B11']
        expected_b11 = EXPECTED_COMMENTS['B11']
        if cell_b11.comment is not None:
            actual_text = normalize_comment(cell_b11.comment.text)
            expected_text = normalize_comment(expected_b11)
            if expected_text in actual_text or actual_text == expected_text:
                print(f"PASS: Component 3 — B11 has benefits enrollment comment: {repr(actual_text[:60])}... (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 3 — B11 comment text mismatch.")
                print(f"  Expected: {repr(expected_b11[:80])}")
                print(f"  Found:    {repr(actual_text[:80])}")
        else:
            print(f"FAIL: Component 3 — B11 has no comment (expected benefits enrollment note)")
    except Exception as e:
        print(f"ERROR: Component 3 — Could not check B11 comment: {e}")

    # Component 4: B15 must have the IT equipment setup comment (0.25 points)
    # This FAILS on initial (no comment) and PASSES on golden (comment present)
    try:
        cell_b15 = ws['B15']
        expected_b15 = EXPECTED_COMMENTS['B15']
        if cell_b15.comment is not None:
            actual_text = normalize_comment(cell_b15.comment.text)
            expected_text = normalize_comment(expected_b15)
            if expected_text in actual_text or actual_text == expected_text:
                print(f"PASS: Component 4 — B15 has IT equipment comment: {repr(actual_text[:60])}... (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 4 — B15 comment text mismatch.")
                print(f"  Expected: {repr(expected_b15[:80])}")
                print(f"  Found:    {repr(actual_text[:80])}")
        else:
            print(f"FAIL: Component 4 — B15 has no comment (expected IT equipment setup note)")
    except Exception as e:
        print(f"ERROR: Component 4 — Could not check B15 comment: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
