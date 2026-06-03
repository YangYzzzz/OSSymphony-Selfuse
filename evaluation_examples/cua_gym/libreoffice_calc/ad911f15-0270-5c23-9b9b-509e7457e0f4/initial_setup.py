"""
Initial Setup: HR Onboarding Checklist — cell comments task
Task ID: calc_hr_onboarding_checklist_comments_029
Domain: libreoffice_calc

Creates the initial Onboarding Checklist spreadsheet with NO comments.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_onboarding_checklist_comments_029'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Onboarding Checklist'

    # ── Row 1: Merged header ─────────────────────────────────────────
    ws.merge_cells('A1:D1')
    ws['A1'] = 'New Employee Onboarding Checklist'
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='1F3864')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36

    # ── Row 2: Column headers ────────────────────────────────────────
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5597', end_color='FF2F5597', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['Done', 'Task', 'Due Date', 'Owner']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    ws.row_dimensions[2].height = 22

    # ── Checklist data ────────────────────────────────────────────────
    # Format: (done, task, due_date, owner)
    # Key rows: 3=I-9 Form, 7=Payroll Direct Deposit, 11=Benefits Enrollment, 15=IT Equipment Setup
    rows_data = [
        # Row 3
        ('☐', 'Complete I-9 Form', '2025-08-05', 'HR Coordinator'),
        # Row 4
        ('☐', 'Sign Employee Handbook Acknowledgement', '2025-08-05', 'HR Coordinator'),
        # Row 5
        ('☐', 'Review Company Policies and Code of Conduct', '2025-08-06', 'HR Coordinator'),
        # Row 6
        ('☐', 'Complete New Hire Tax Forms (W-4, State Withholding)', '2025-08-06', 'Payroll'),
        # Row 7
        ('☐', 'Set Up Payroll Direct Deposit', '2025-08-08', 'Payroll'),
        # Row 8
        ('☐', 'Obtain Employee ID Badge', '2025-08-07', 'Facilities'),
        # Row 9
        ('☐', 'Complete Security Awareness Training', '2025-08-14', 'IT'),
        # Row 10
        ('☐', 'Complete Workplace Safety Orientation', '2025-08-09', 'Safety Officer'),
        # Row 11
        ('☐', 'Complete Benefits Enrollment', '2025-09-01', 'Benefits Admin'),
        # Row 12
        ('☐', 'Set Up Company Email Account', '2025-08-04', 'IT'),
        # Row 13
        ('☐', 'Schedule 30-Day Check-In with Manager', '2025-09-01', 'Manager'),
        # Row 14
        ('☐', 'Complete Compliance Training (HIPAA/SOX if applicable)', '2025-08-21', 'Compliance'),
        # Row 15
        ('☐', 'IT Equipment Setup', '2025-07-28', 'IT'),
        # Row 16
        ('☐', 'Set Up VPN Access and Two-Factor Authentication', '2025-08-04', 'IT'),
        # Row 17
        ('☐', 'Attend New Employee Orientation Session', '2025-08-04', 'HR Coordinator'),
        # Row 18
        ('☐', 'Meet with Direct Manager for Role Overview', '2025-08-04', 'Manager'),
        # Row 19
        ('☐', 'Enroll in Required Professional Development Courses', '2025-08-29', 'L&D'),
    ]

    row_fill_even = PatternFill(start_color='FFDCE6F1', end_color='FFDCE6F1', fill_type='solid')
    row_fill_odd = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
    data_font = Font(name='Calibri', size=11)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for i, (done, task, due_date, owner) in enumerate(rows_data):
        row_num = i + 3  # starts at row 3
        fill = row_fill_even if i % 2 == 0 else row_fill_odd

        # Column A: Checkbox symbol
        a = ws.cell(row=row_num, column=1, value=done)
        a.font = Font(name='Calibri', size=13)
        a.alignment = center_align
        a.fill = fill
        a.border = border

        # Column B: Task name
        b = ws.cell(row=row_num, column=2, value=task)
        b.font = data_font
        b.alignment = left_align
        b.fill = fill
        b.border = border

        # Column C: Due date
        c = ws.cell(row=row_num, column=3, value=due_date)
        c.font = data_font
        c.alignment = center_align
        c.fill = fill
        c.border = border

        # Column D: Owner
        d = ws.cell(row=row_num, column=4, value=owner)
        d.font = data_font
        d.alignment = center_align
        d.fill = fill
        d.border = border

        ws.row_dimensions[row_num].height = 22

    # ── Column widths ────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 48
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 20

    # ── Freeze pane below header rows ───────────────────────────────
    ws.freeze_panes = 'A3'

    # ── NO comments on any cells ─────────────────────────────────────
    # (This is the initial state; comments will be added by the agent)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: Onboarding Checklist')
    print('Rows: 2 header rows + 17 data rows (rows 3-19)')
    print('Comments: NONE (task requires agent to add them)')


create_initial()
