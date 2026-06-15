"""
Initial Setup: Create workbook with empty 'Links' sheet and external Targets.xlsx
Task ID: calc_mcp_062
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_062'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'
TARGETS_DIR = f'{WORKDIR}/Documents'
TARGETS_FILE = f'{TARGETS_DIR}/Targets.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    # --- Create the main workbook with empty Links sheet ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Links'
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # --- Create the external reference workbook Targets.xlsx ---
    os.makedirs(TARGETS_DIR, exist_ok=True)

    twb = openpyxl.Workbook()

    # Sheet 'Goals' with B3 = 500000
    ws_goals = twb.active
    ws_goals.title = 'Goals'
    # Add some context data to make it realistic
    ws_goals.cell(row=1, column=1, value='Category')
    ws_goals.cell(row=1, column=2, value='Target')
    ws_goals.cell(row=1, column=3, value='Status')
    ws_goals.cell(row=2, column=1, value='Q1 Revenue')
    ws_goals.cell(row=2, column=2, value=250000)
    ws_goals.cell(row=2, column=3, value='Achieved')
    ws_goals.cell(row=3, column=1, value='Annual Revenue')
    ws_goals.cell(row=3, column=2, value=500000)
    ws_goals.cell(row=3, column=3, value='In Progress')
    ws_goals.cell(row=4, column=1, value='Customer Acquisition')
    ws_goals.cell(row=4, column=2, value=1200)
    ws_goals.cell(row=4, column=3, value='On Track')
    ws_goals.cell(row=5, column=1, value='Market Share')
    ws_goals.cell(row=5, column=2, value=15.5)
    ws_goals.cell(row=5, column=3, value='Behind')

    # Sheet 'Metrics' with D10 = 92.5
    ws_metrics = twb.create_sheet('Metrics')
    ws_metrics.cell(row=1, column=1, value='Metric')
    ws_metrics.cell(row=1, column=2, value='Q1')
    ws_metrics.cell(row=1, column=3, value='Q2')
    ws_metrics.cell(row=1, column=4, value='Q3')
    ws_metrics.cell(row=1, column=5, value='Q4')
    metrics_data = [
        ['Customer Satisfaction', 88.2, 89.1, 90.3, 91.0],
        ['Employee Engagement', 76.5, 78.0, 79.2, 80.1],
        ['Net Promoter Score', 45, 47, 49, 52],
        ['Conversion Rate', 3.2, 3.5, 3.8, 4.1],
        ['Churn Rate', 2.1, 1.9, 1.7, 1.5],
        ['Support Tickets Resolved', 1250, 1340, 1410, 1520],
        ['Avg Response Time (hrs)', 4.2, 3.8, 3.5, 3.1],
        ['Revenue Growth (%)', 12.3, 14.1, 15.8, 17.2],
        ['Profit Margin (%)', 85.0, 88.5, 92.5, 95.0],
    ]
    for r, row_data in enumerate(metrics_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_metrics.cell(row=r, column=c, value=val)

    twb.save(TARGETS_FILE)
    print(f'External reference file created: {TARGETS_FILE}')

    # GUI-ready startup: open the main workbook in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
