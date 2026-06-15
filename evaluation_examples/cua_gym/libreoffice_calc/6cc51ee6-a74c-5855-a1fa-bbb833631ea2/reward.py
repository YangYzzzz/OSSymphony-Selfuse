"""
Reward Script: External workbook references in LibreOffice Calc
Task ID: calc_mcp_062
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): A1 contains an external reference formula pointing to Targets.xlsx Goals!B3
  Component 2 (0.5): A2 contains an external reference formula pointing to Targets.xlsx Metrics!D10
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_062'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice document."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def normalize_formula(formula_str):
    """
    Normalize an external reference formula for comparison.
    Strips spaces, lowercases, removes optional quotes/brackets variations.
    Returns a tuple: (file_path_lower, sheet_name_lower, cell_ref_upper) or None.
    """
    if not isinstance(formula_str, str):
        return None
    f = formula_str.strip()
    if not f.startswith('='):
        return None

    # Remove leading '='
    f = f[1:]

    # Pattern for LibreOffice-style external references:
    #   '[/path/to/file.xlsx]SheetName'.CellRef
    #   OR '[/path/to/file.xlsx]SheetName'.CellRef (with quotes around sheet part)
    #   OR various quote patterns
    # We need to extract: file path, sheet name, cell reference

    # Try multiple patterns that LibreOffice / openpyxl might produce
    patterns = [
        # Pattern: '[filepath]sheet'.cell  (LO style with quotes around whole ref)
        r"""^'?\[([^\]]+)\]([^']+)'?\.([A-Za-z]+\d+)$""",
        # Pattern: '[filepath]sheet'!cell  (Excel-like with !)
        r"""^'?\[([^\]]+)\]([^']+)'?!([A-Za-z]+\d+)$""",
        # Pattern: 'filepath'#$sheet.cell  (another LO variant)
        r"""^'([^']+)'#\$?([^.]+)\.([A-Za-z]+\d+)$""",
    ]

    for pat in patterns:
        m = re.match(pat, f)
        if m:
            fpath = m.group(1).strip().lower()
            sheet = m.group(2).strip().strip("'").lower()
            cell = m.group(3).strip().upper()
            return (fpath, sheet, cell)

    return None


def check_external_ref(formula_str, expected_file_substring, expected_sheet, expected_cell):
    """
    Check if a formula references the expected external file, sheet, and cell.
    Returns True if the reference matches (case-insensitive for path/sheet, case-insensitive for cell).
    """
    parsed = normalize_formula(formula_str)
    if parsed is None:
        return False

    fpath, sheet, cell = parsed
    file_ok = expected_file_substring.lower() in fpath
    sheet_ok = sheet == expected_sheet.lower()
    cell_ok = cell == expected_cell.upper()
    return file_ok and sheet_ok and cell_ok


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Links' sheet must exist
    if 'Links' not in wb.sheetnames:
        print(f"FAIL: 'Links' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Links']

    # Component 1: A1 contains external reference to Targets.xlsx Goals!B3 (0.5 points)
    try:
        a1_val = ws['A1'].value
        if a1_val is not None and isinstance(a1_val, str) and '=' in a1_val[:2]:
            if check_external_ref(a1_val, 'targets.xlsx', 'Goals', 'B3'):
                print(f"PASS: Component 1 — A1 has correct external ref to Targets.xlsx Goals!B3 (0.5 pts)")
                print(f"  Formula: {a1_val}")
                total_score += 0.5
            else:
                print(f"FAIL: Component 1 — A1 formula does not match expected external ref")
                print(f"  Found: {a1_val}")
                print(f"  Expected: reference to Targets.xlsx, sheet Goals, cell B3")
        else:
            print(f"FAIL: Component 1 — A1 is not a formula. Value: {a1_val}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: A2 contains external reference to Targets.xlsx Metrics!D10 (0.5 points)
    try:
        a2_val = ws['A2'].value
        if a2_val is not None and isinstance(a2_val, str) and '=' in a2_val[:2]:
            if check_external_ref(a2_val, 'targets.xlsx', 'Metrics', 'D10'):
                print(f"PASS: Component 2 — A2 has correct external ref to Targets.xlsx Metrics!D10 (0.5 pts)")
                print(f"  Formula: {a2_val}")
                total_score += 0.5
            else:
                print(f"FAIL: Component 2 — A2 formula does not match expected external ref")
                print(f"  Found: {a2_val}")
                print(f"  Expected: reference to Targets.xlsx, sheet Metrics, cell D10")
        else:
            print(f"FAIL: Component 2 — A2 is not a formula. Value: {a2_val}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist any unsaved state
persist_app_state("libreoffice_calc")

# Run verification
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
