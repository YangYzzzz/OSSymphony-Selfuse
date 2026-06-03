"""
Reward Script: Re-import CSV data with semicolon delimiter to split into 5 columns
Task ID: calc_tbl_032
Domain: libreoffice_calc

Scoring Rubric:
  Component 1 (0.3): Data spans 5 columns (max_column >= 5)
  Component 2 (0.3): Header row correctly split into 5 separate headers
  Component 3 (0.2): All data rows have non-None values in columns B-E
  Component 4 (0.2): Salary column (D) contains numeric values, not strings with semicolons
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_032'

def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Component 1: Data spans 5 columns (0.3 points)
    # Initial env has max_column == 1 (everything in col A).
    # Golden env should have max_column >= 5.
    try:
        max_col = ws.max_column
        if max_col >= 5:
            print(f"PASS: Component 1 — max_column={max_col} >= 5 (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — max_column={max_col}, expected >= 5")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Header row correctly split (0.3 points)
    # Initial env has A1='First Name;Last Name;Department;Salary;Year' (single cell).
    # Golden env should have A1='First Name', B1='Last Name', C1='Department', D1='Salary', E1='Year'.
    try:
        expected_headers = ['First Name', 'Last Name', 'Department', 'Salary', 'Year']
        actual_headers = []
        for col in range(1, 6):
            val = ws.cell(row=1, column=col).value
            actual_headers.append(str(val).strip() if val is not None else None)

        matches = sum(1 for a, e in zip(actual_headers, expected_headers) if a == e)
        if matches == 5:
            print(f"PASS: Component 2 — All 5 headers correct: {actual_headers} (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — Headers: {actual_headers}, expected: {expected_headers} ({matches}/5 match)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: All data rows have values in columns B through E (0.2 points)
    # Initial env has B2:E16 all None (data only in col A).
    # Golden env should have non-None values in B-E for all 15 data rows.
    try:
        total_data_rows = 0
        rows_with_all_cols = 0
        for row_num in range(2, ws.max_row + 1):
            a_val = ws.cell(row=row_num, column=1).value
            if a_val is None:
                continue
            total_data_rows += 1
            b_val = ws.cell(row=row_num, column=2).value
            c_val = ws.cell(row=row_num, column=3).value
            d_val = ws.cell(row=row_num, column=4).value
            e_val = ws.cell(row=row_num, column=5).value
            if b_val is not None and c_val is not None and d_val is not None and e_val is not None:
                rows_with_all_cols += 1

        if total_data_rows > 0 and rows_with_all_cols == total_data_rows:
            print(f"PASS: Component 3 — All {total_data_rows} data rows have values in B-E (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — {rows_with_all_cols}/{total_data_rows} rows have all columns filled")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Salary column (D) contains numeric values (0.2 points)
    # Initial env col D is all None. Golden env should have integers like 50000, 85000 etc.
    # Also verify col A no longer contains semicolons (data was split).
    try:
        numeric_count = 0
        no_semicolons_in_a = 0
        data_rows = 0
        for row_num in range(2, ws.max_row + 1):
            a_val = ws.cell(row=row_num, column=1).value
            if a_val is None:
                continue
            data_rows += 1

            # Check col A does NOT contain semicolons (was properly split)
            if ';' not in str(a_val):
                no_semicolons_in_a += 1

            # Check col D is numeric
            d_val = ws.cell(row=row_num, column=4).value
            if isinstance(d_val, (int, float)):
                numeric_count += 1

        if data_rows > 0 and numeric_count == data_rows and no_semicolons_in_a == data_rows:
            print(f"PASS: Component 4 — All {data_rows} rows: D is numeric, A has no semicolons (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 4 — numeric D: {numeric_count}/{data_rows}, no-semicolons in A: {no_semicolons_in_a}/{data_rows}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
