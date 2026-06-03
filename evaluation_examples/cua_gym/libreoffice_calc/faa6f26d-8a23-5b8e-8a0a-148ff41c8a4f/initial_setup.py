"""
Initial Setup: CSV data imported with wrong delimiter - all in column A with semicolons
Task ID: calc_tbl_032
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_032'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Simulate a bad CSV import: all data crammed into column A with semicolons
    # Row 1: header row (also semicolon-separated in column A)
    ws.cell(row=1, column=1, value="First Name;Last Name;Department;Salary;Year")

    # Rows 2+: realistic employee data, all in column A as semicolon-separated strings
    employees = [
        "John;Smith;Sales;50000;2020",
        "Sarah;Chen;Engineering;85000;2021",
        "Marcus;Johnson;Marketing;72000;2019",
        "Emily;Davis;Finance;68000;2022",
        "Robert;Wilson;Engineering;91000;2018",
        "Lisa;Anderson;Sales;53000;2021",
        "David;Martinez;Finance;75000;2020",
        "Jennifer;Taylor;Marketing;64000;2023",
        "Michael;Brown;Engineering;88000;2019",
        "Amanda;Garcia;Sales;57000;2022",
        "Christopher;Lee;Finance;71000;2020",
        "Jessica;Thomas;Marketing;66000;2021",
        "Daniel;Jackson;Engineering;93000;2018",
        "Rachel;White;Sales;55000;2023",
        "Kevin;Harris;Finance;79000;2019",
    ]

    for r, emp in enumerate(employees, 2):
        ws.cell(row=r, column=1, value=emp)

    # Set column A wide enough to show the full semicolon-separated strings
    ws.column_dimensions["A"].width = 55

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
