"""
Initial Setup: Financial model workbook with hardcoded tax/discount/inflation rates
Task ID: calc_gen_namedranges_050
Domain: libreoffice_calc

Creates a workbook with 5 sheets (Assumptions, Income, Balance, Cashflow, Valuation)
where financial parameters (tax rate 0.21, discount rate 0.10, inflation rate 0.025)
are hardcoded in formulas — NOT yet using named ranges.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_namedranges_050'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # ---------------------------------------------------------------
    # Sheet 1: Assumptions
    # ---------------------------------------------------------------
    ws_assump = wb.active
    ws_assump.title = 'Assumptions'

    header_font = Font(bold=True, size=12)
    ws_assump['A1'] = 'Financial Model Assumptions'
    ws_assump['A1'].font = Font(bold=True, size=14)
    ws_assump.column_dimensions['A'].width = 22
    ws_assump.column_dimensions['B'].width = 16
    ws_assump.column_dimensions['C'].width = 28

    ws_assump['A2'] = 'Tax Rate'
    ws_assump['B2'] = 0.21
    ws_assump['B2'].number_format = '0.00%'
    ws_assump['C2'] = 'Corporate income tax rate'

    ws_assump['A3'] = 'Discount Rate'
    ws_assump['B3'] = 0.10
    ws_assump['B3'].number_format = '0.00%'
    ws_assump['C3'] = 'WACC / hurdle rate'

    ws_assump['A4'] = 'Inflation Rate'
    ws_assump['B4'] = 0.025
    ws_assump['B4'].number_format = '0.00%'
    ws_assump['C4'] = 'Annual inflation assumption'

    ws_assump['A6'] = 'Fiscal Year'
    ws_assump['B6'] = 2025
    ws_assump['C6'] = 'Base year for projections'

    ws_assump['A7'] = 'Projection Years'
    ws_assump['B7'] = 5
    ws_assump['C7'] = 'Number of forecast years'

    for row in [2, 3, 4, 6, 7]:
        ws_assump[f'A{row}'].font = header_font

    # ---------------------------------------------------------------
    # Sheet 2: Income
    # ---------------------------------------------------------------
    ws_inc = wb.create_sheet('Income')
    ws_inc['A1'] = 'Income Statement (5-Year Forecast)'
    ws_inc['A1'].font = Font(bold=True, size=13)

    years = [2025, 2026, 2027, 2028, 2029]
    ws_inc['A2'] = 'Item'
    for i, yr in enumerate(years, 2):
        ws_inc.cell(row=2, column=i, value=yr)
        ws_inc.cell(row=2, column=i).font = Font(bold=True)

    inc_rows = [
        ('Revenue',           [4200000, 4452000, 4719120, 5002067, 5302191]),
        ('COGS',              [2100000, 2226000, 2359560, 2501034, 2651096]),
        ('Gross Profit',      [None]*5),  # formula
        ('Operating Expenses',[840000,  891120,  945187, 1002414, 1060782]),
        ('EBIT',              [None]*5),  # formula
        ('Interest Expense',  [84000,   79800,   75600,   71400,   67200]),
        ('EBT',               [None]*5),  # formula
        ('Tax Expense',       [None]*5),  # formula = EBT * 0.21
        ('Net Income',        [None]*5),  # formula
    ]
    row_labels = [r[0] for r in inc_rows]
    row_data   = [r[1] for r in inc_rows]

    row_map = {}
    for ri, (label, vals) in enumerate(inc_rows, 3):
        ws_inc.cell(row=ri, column=1, value=label)
        ws_inc.cell(row=ri, column=1).font = Font(bold=(label in ('Gross Profit','EBIT','EBT','Net Income','Tax Expense')))
        row_map[label] = ri
        for ci, val in enumerate(vals, 2):
            if val is not None:
                ws_inc.cell(row=ri, column=ci, value=val)
                ws_inc.cell(row=ri, column=ci).number_format = '#,##0'

    col_letters = ['B','C','D','E','F']
    for ci, cl in enumerate(col_letters):
        gp_row  = row_map['Gross Profit']
        ebit_row = row_map['EBIT']
        ebt_row  = row_map['EBT']
        tax_row  = row_map['Tax Expense']
        ni_row   = row_map['Net Income']
        rev_row  = row_map['Revenue']
        cogs_row = row_map['COGS']
        oe_row   = row_map['Operating Expenses']
        ie_row   = row_map['Interest Expense']

        # Gross Profit = Revenue - COGS
        ws_inc[f'{cl}{gp_row}'] = f'={cl}{rev_row}-{cl}{cogs_row}'
        ws_inc[f'{cl}{gp_row}'].number_format = '#,##0'
        # EBIT = Gross Profit - Operating Expenses
        ws_inc[f'{cl}{ebit_row}'] = f'={cl}{gp_row}-{cl}{oe_row}'
        ws_inc[f'{cl}{ebit_row}'].number_format = '#,##0'
        # EBT = EBIT - Interest Expense
        ws_inc[f'{cl}{ebt_row}'] = f'={cl}{ebit_row}-{cl}{ie_row}'
        ws_inc[f'{cl}{ebt_row}'].number_format = '#,##0'
        # Tax Expense = EBT * 0.21  (HARDCODED - to be replaced with named range)
        ws_inc[f'{cl}{tax_row}'] = f'={cl}{ebt_row}*0.21'
        ws_inc[f'{cl}{tax_row}'].number_format = '#,##0'
        # Net Income = EBT - Tax Expense
        ws_inc[f'{cl}{ni_row}'] = f'={cl}{ebt_row}-{cl}{tax_row}'
        ws_inc[f'{cl}{ni_row}'].number_format = '#,##0'

    ws_inc.column_dimensions['A'].width = 22
    for col in ['B','C','D','E','F']:
        ws_inc.column_dimensions[col].width = 14

    # ---------------------------------------------------------------
    # Sheet 3: Balance
    # ---------------------------------------------------------------
    ws_bal = wb.create_sheet('Balance')
    ws_bal['A1'] = 'Balance Sheet Metrics'
    ws_bal['A1'].font = Font(bold=True, size=13)

    ws_bal['A2'] = 'Item'
    for i, yr in enumerate(years, 2):
        ws_bal.cell(row=2, column=i, value=yr)
        ws_bal.cell(row=2, column=i).font = Font(bold=True)

    # Deferred tax liability uses tax rate; real assets inflated
    assets_base = [1500000, 1600000, 1700000, 1820000, 1950000]
    liab_base   = [600000,  620000,  640000,  660000,  680000]
    ws_bal['A3'] = 'Total Assets'
    ws_bal['A4'] = 'Inflation-Adjusted Assets'
    ws_bal['A5'] = 'Total Liabilities'
    ws_bal['A6'] = 'Deferred Tax Liability'
    ws_bal['A7'] = 'Shareholders Equity'

    for ci, cl in enumerate(col_letters):
        yr_idx = ci
        ws_bal[f'{cl}3'] = assets_base[yr_idx]
        ws_bal[f'{cl}3'].number_format = '#,##0'
        # Inflation-adjusted: Assets * (1 + 0.025)  [hardcoded]
        ws_bal[f'{cl}4'] = f'={cl}3*(1+0.025)'
        ws_bal[f'{cl}4'].number_format = '#,##0'
        ws_bal[f'{cl}5'] = liab_base[yr_idx]
        ws_bal[f'{cl}5'].number_format = '#,##0'
        # Deferred Tax = Liabilities * 0.21  [hardcoded]
        ws_bal[f'{cl}6'] = f'={cl}5*0.21'
        ws_bal[f'{cl}6'].number_format = '#,##0'
        # Equity = Assets - Liabilities
        ws_bal[f'{cl}7'] = f'={cl}3-{cl}5'
        ws_bal[f'{cl}7'].number_format = '#,##0'

    ws_bal.column_dimensions['A'].width = 28
    for col in ['B','C','D','E','F']:
        ws_bal.column_dimensions[col].width = 14

    # ---------------------------------------------------------------
    # Sheet 4: Cashflow
    # ---------------------------------------------------------------
    ws_cf = wb.create_sheet('Cashflow')
    ws_cf['A1'] = 'Cash Flow Statement'
    ws_cf['A1'].font = Font(bold=True, size=13)

    ws_cf['A2'] = 'Item'
    for i, yr in enumerate(years, 2):
        ws_cf.cell(row=2, column=i, value=yr)
        ws_cf.cell(row=2, column=i).font = Font(bold=True)

    cf_operating = [1200000, 1272000, 1348320, 1429219, 1514972]
    cf_investing  = [-300000, -318000, -337080, -357305, -378743]
    cf_tax_paid   = None  # computed from operating * 0.21

    ws_cf['A3'] = 'Operating Cash Flow'
    ws_cf['A4'] = 'Taxes Paid'        # = operating * 0.21 (hardcoded)
    ws_cf['A5'] = 'Investing Activities'
    ws_cf['A6'] = 'Inflation Adjustment'   # = investing * 0.025
    ws_cf['A7'] = 'Discount Factor'         # = 1/(1+0.10)^n
    ws_cf['A8'] = 'PV of Cash Flow'         # = operating * discount
    ws_cf['A9'] = 'Net Free Cash Flow'

    for ci, cl in enumerate(col_letters):
        yr_idx = ci
        ws_cf[f'{cl}3'] = cf_operating[yr_idx]
        ws_cf[f'{cl}3'].number_format = '#,##0'
        # Tax paid = operating * 0.21 [hardcoded]
        ws_cf[f'{cl}4'] = f'={cl}3*0.21'
        ws_cf[f'{cl}4'].number_format = '#,##0'
        ws_cf[f'{cl}5'] = cf_investing[yr_idx]
        ws_cf[f'{cl}5'].number_format = '#,##0'
        # Inflation adjustment = investing * 0.025 [hardcoded]
        ws_cf[f'{cl}6'] = f'={cl}5*0.025'
        ws_cf[f'{cl}6'].number_format = '#,##0'
        # Discount factor = 1/(1+0.10)^n [hardcoded], n = yr_idx+1
        n = yr_idx + 1
        ws_cf[f'{cl}7'] = f'=1/(1+0.10)^{n}'
        ws_cf[f'{cl}7'].number_format = '0.0000'
        # PV of Cash Flow = operating * discount factor
        ws_cf[f'{cl}8'] = f'={cl}3*{cl}7'
        ws_cf[f'{cl}8'].number_format = '#,##0'
        # Net FCF = operating + investing
        ws_cf[f'{cl}9'] = f'={cl}3+{cl}5'
        ws_cf[f'{cl}9'].number_format = '#,##0'

    ws_cf.column_dimensions['A'].width = 24
    for col in ['B','C','D','E','F']:
        ws_cf.column_dimensions[col].width = 14

    # ---------------------------------------------------------------
    # Sheet 5: Valuation
    # ---------------------------------------------------------------
    ws_val = wb.create_sheet('Valuation')
    ws_val['A1'] = 'Valuation Summary'
    ws_val['A1'].font = Font(bold=True, size=13)

    ws_val['A3'] = 'Valuation Parameter'
    ws_val['B3'] = 'Value'
    ws_val['C3'] = 'Notes'
    ws_val['A3'].font = Font(bold=True)
    ws_val['B3'].font = Font(bold=True)
    ws_val['C3'].font = Font(bold=True)

    ws_val['A4'] = 'WACC Used'
    ws_val['B4'] = 0.10                   # hardcoded value used in formula
    ws_val['B4'].number_format = '0.00%'
    ws_val['C4'] = 'Discount rate applied in DCF'

    ws_val['A5'] = 'Terminal Growth Rate'
    ws_val['B5'] = 0.025                  # = inflation rate, hardcoded
    ws_val['B5'].number_format = '0.00%'
    ws_val['C5'] = 'Long-run nominal growth'

    ws_val['A6'] = 'Effective Tax Rate'
    ws_val['B6'] = 0.21
    ws_val['B6'].number_format = '0.00%'
    ws_val['C6'] = 'Used in NOPAT calc'

    ws_val['A8'] = 'Year'
    ws_val['B8'] = 'Free Cash Flow'
    ws_val['C8'] = 'Discount Factor'
    ws_val['D8'] = 'PV of FCF'
    for col in ['A','B','C','D']:
        ws_val[f'{col}8'].font = Font(bold=True)

    fcf_vals = [900000, 954000, 1011240, 1071914, 1136229]
    for ri, (yr, fcf) in enumerate(zip(years, fcf_vals), 9):
        n = ri - 8
        ws_val[f'A{ri}'] = yr
        ws_val[f'B{ri}'] = fcf
        ws_val[f'B{ri}'].number_format = '#,##0'
        # Discount factor = 1/(1+0.10)^n [hardcoded]
        ws_val[f'C{ri}'] = f'=1/(1+0.10)^{n}'
        ws_val[f'C{ri}'].number_format = '0.0000'
        # PV
        ws_val[f'D{ri}'] = f'=B{ri}*C{ri}'
        ws_val[f'D{ri}'].number_format = '#,##0'

    # Terminal value row
    tv_row = 14
    ws_val[f'A{tv_row}'] = 'Terminal Value'
    # TV = FCF_last * (1+0.025) / (0.10 - 0.025)  [both hardcoded]
    ws_val[f'B{tv_row}'] = f'=B13*(1+0.025)/(0.10-0.025)'
    ws_val[f'B{tv_row}'].number_format = '#,##0'
    ws_val[f'C{tv_row}'] = f'=1/(1+0.10)^5'
    ws_val[f'C{tv_row}'].number_format = '0.0000'
    ws_val[f'D{tv_row}'] = f'=B{tv_row}*C{tv_row}'
    ws_val[f'D{tv_row}'].number_format = '#,##0'

    # NOPAT row uses 0.21
    nopat_row = 16
    ws_val[f'A{nopat_row}'] = 'NOPAT (Year 1)'
    ebit_year1 = 1176000
    ws_val[f'B{nopat_row}'] = f'={ebit_year1}*(1-0.21)'
    ws_val[f'B{nopat_row}'].number_format = '#,##0'
    ws_val[f'C{nopat_row}'] = 'EBIT*(1-TaxRate)'

    ws_val['A18'] = 'Enterprise Value (Sum of PV FCF + Terminal PV)'
    ws_val['B18'] = f'=SUM(D9:D{tv_row})'
    ws_val['B18'].number_format = '#,##0'
    ws_val['A18'].font = Font(bold=True)
    ws_val['B18'].font = Font(bold=True)

    ws_val.column_dimensions['A'].width = 44
    ws_val.column_dimensions['B'].width = 18
    ws_val.column_dimensions['C'].width = 16
    ws_val.column_dimensions['D'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheets: {wb.sheetnames}')
    print('Hardcoded rates in formulas (0.21, 0.10, 0.025) — no named ranges defined')


create_initial()
