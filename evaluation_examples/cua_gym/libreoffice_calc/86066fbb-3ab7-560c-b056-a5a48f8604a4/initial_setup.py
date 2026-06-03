"""
Initial Setup: Add custom asymmetric error bars to chart
Task ID: calc_gg2_025
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.chart import BarChart, Reference

WORKDIR = '/home/user'
TASK_ID = 'calc_gg2_025'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Projections ---
    ws = wb.active
    ws.title = 'Projections'

    # Headers
    headers = ['Month', 'Forecast', 'Actual', 'Upper CI Offset', 'Lower CI Offset']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Monthly data: forecast revenue, actual, upper/lower confidence interval offsets
    data = [
        ['January',   125000, 121500, 18000, 12000],
        ['February',  132000, 129800, 21000, 14500],
        ['March',     148000, 150200, 16500, 11000],
        ['April',     155000, 152300, 24000, 16000],
        ['May',       163000, 167500, 19500, 13500],
        ['June',      171000, 168200, 22000, 15000],
        ['July',      158000, 155800, 20500, 14000],
        ['August',    165000, 162100, 17500, 12500],
        ['September', 178000, 180500, 25000, 17000],
        ['October',   185000, 182700, 23000, 15500],
        ['November',  192000, 195300, 19000, 13000],
        ['December',  205000, 201800, 26000, 18000],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18

    # Number format for currency columns
    for r in range(2, 14):
        for c in [2, 3, 4, 5]:
            ws.cell(row=r, column=c).number_format = '#,##0'

    # --- Embedded Bar Chart (Forecast by Month, NO error bars) ---
    chart = BarChart()
    chart.type = "col"
    chart.title = "Revenue Forecast by Month"
    chart.y_axis.title = "Revenue ($)"
    chart.x_axis.title = "Month"
    chart.style = 10

    # Forecast data series (B1:B13 with header)
    forecast_data = Reference(ws, min_col=2, min_row=1, max_row=13)
    categories = Reference(ws, min_col=1, min_row=2, max_row=13)
    chart.add_data(forecast_data, titles_from_data=True)
    chart.set_categories(categories)

    chart.width = 20
    chart.height = 12

    ws.add_chart(chart, "G2")

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
