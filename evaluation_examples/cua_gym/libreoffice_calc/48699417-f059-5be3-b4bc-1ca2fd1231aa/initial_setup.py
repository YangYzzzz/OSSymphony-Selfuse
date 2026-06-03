"""
Initial Setup: Create research data spreadsheet with 8 columns and 300 rows on 'Data' sheet.
No header or footer is configured — the researcher needs to add page numbering.
Task ID: calc_gg1_037
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_037'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Data'

    # Headers for research data
    headers = [
        'Subject ID', 'Name', 'Age', 'Treatment Group',
        'Baseline Score', 'Week 4 Score', 'Week 8 Score', 'Adverse Events'
    ]
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # Set column widths
    col_widths = [12, 22, 8, 18, 16, 16, 16, 16]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[chr(65 + i)].width = w

    # Generate 300 rows of realistic research data
    first_names = [
        'Sarah', 'Marcus', 'Yuki', 'Elena', 'Rajesh', 'Amara', 'Wei',
        'Fatima', 'Lucas', 'Priya', 'Dmitri', 'Mei', 'Carlos', 'Aisha',
        'James', 'Nora', 'Kenji', 'Olivia', 'Hassan', 'Ling', 'Sofia',
        'David', 'Anya', 'Tomás', 'Ingrid', 'Omar', 'Hana', 'Erik',
        'Zara', 'Ravi'
    ]
    last_names = [
        'Chen', 'Johnson', 'Tanaka', 'Petrov', 'Sharma', 'Okafor', 'Zhang',
        'Al-Hassan', 'Müller', 'Gupta', 'Ivanov', 'Liu', 'Rodriguez', 'Patel',
        'Williams', 'Svensson', 'Yamamoto', 'Martin', 'Abbas', 'Wang', 'Garcia',
        'Kim', 'Novak', 'Hernandez', 'Larsson', 'Ibrahim', 'Suzuki', 'Jensen',
        'Bakshi', 'Kowalski'
    ]
    groups = ['Placebo', 'Treatment A', 'Treatment B', 'Treatment C']
    adverse_options = [
        'None', 'None', 'None', 'None', 'None',  # weighted toward none
        'Mild headache', 'Nausea', 'Fatigue', 'Dizziness',
        'Insomnia', 'Mild rash', 'Joint pain', 'Dry mouth'
    ]

    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r in range(2, 302):  # rows 2 to 301 (300 data rows)
        subject_id = f'SUBJ-{r - 1:04d}'
        name = f'{random.choice(first_names)} {random.choice(last_names)}'
        age = random.randint(22, 78)
        group = random.choice(groups)
        baseline = round(random.uniform(35.0, 95.0), 1)
        week4 = round(baseline + random.uniform(-15.0, 25.0), 1)
        week8 = round(week4 + random.uniform(-10.0, 20.0), 1)
        adverse = random.choice(adverse_options)

        row_data = [subject_id, name, age, group, baseline, week4, week8, adverse]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = data_border
            if c in (5, 6, 7):
                cell.number_format = '0.0'

    # Ensure NO header or footer is set (default is no header/footer)
    # Explicitly confirm oddFooter and oddHeader are not set
    # (openpyxl default is no header/footer, so nothing to do)

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
