"""
Initial Setup: Apply custom phone number format to contacts spreadsheet
Task ID: calc_lf_068
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_068'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Contacts ---
    ws = wb.active
    ws.title = 'Contacts'

    # Headers with light styling
    headers = ['Name', 'Phone']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
    header_alignment = Alignment(horizontal='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Data rows - realistic contacts with 10-digit phone numbers
    # Phone numbers stored as plain integers with NO custom format (General)
    data = [
        ['Alice Martinez', 5551234567],
        ['Bob Richardson', 2129876543],
        ['Carol Nakamura', 3105551212],
        ['David Okonkwo', 4155559832],
        ['Elena Petrova', 7185553741],
        ['Frank Sullivan', 6175558294],
        ['Grace Kim', 8585551063],
        ['Hector Ramirez', 9045557821],
        ['Irene Chang', 5035554490],
        ['James Whitfield', 2025556173],
        ['Karen Johansson', 3125559045],
        ['Liam Patel', 6465553287],
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        phone_cell = ws.cell(row=r, column=2, value=row_data[1])
        # Ensure phone numbers are stored as numbers with General format (default)

    # Set column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18

    # --- Sheet 2: Departments ---
    ws2 = wb.create_sheet('Departments')
    dept_headers = ['Department', 'Head', 'Office Phone', 'Extension']
    for col, h in enumerate(dept_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    dept_data = [
        ['Engineering', 'Sarah Chen', 5551001000, 201],
        ['Marketing', 'Marcus Johnson', 5551002000, 302],
        ['Finance', 'Priya Sharma', 5551003000, 403],
        ['Human Resources', 'Tom Bradley', 5551004000, 504],
        ['Operations', 'Linda Wu', 5551005000, 605],
    ]

    for r, row_data in enumerate(dept_data, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
