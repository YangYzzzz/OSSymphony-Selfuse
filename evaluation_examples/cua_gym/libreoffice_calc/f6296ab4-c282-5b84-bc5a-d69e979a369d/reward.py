"""
Reward Script: Apply custom phone number format (000) 000-0000 to cells B2:B4
Task ID: calc_lf_068
Domain: libreoffice_calc
Scoring:
  - Component 1: B2 has phone format and numeric value (0.35 pts)
  - Component 2: B3 has phone format and numeric value (0.35 pts)
  - Component 3: B4 has phone format and numeric value (0.30 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_068'
EXPECTED_FORMAT = '(000) 000-0000'

# Expected phone numbers (must remain as numeric values, not strings)
EXPECTED_VALUES = {
    'B2': 5551234567,
    'B3': 2129876543,
    'B4': 3105551212,
}

WEIGHTS = {
    'B2': 0.35,
    'B3': 0.35,
    'B4': 0.30,
}


def verify_task(file_path):
    """
    Verify that cells B2:B4 in the 'Contacts' sheet have the custom phone
    number format '(000) 000-0000' applied, and that their values remain
    as numeric integers (not converted to strings).
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Contacts' sheet must exist
    if 'Contacts' not in wb.sheetnames:
        print("FAIL: 'Contacts' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Contacts']

    for i, (coord, expected_val) in enumerate(EXPECTED_VALUES.items(), 1):
        weight = WEIGHTS[coord]
        # Component i: Check phone format and numeric value for cell
        try:
            cell = ws[coord]
            cell_val = cell.value
            cell_fmt = cell.number_format

            # Check 1: number_format matches the phone format
            fmt_ok = (cell_fmt == EXPECTED_FORMAT)

            # Check 2: value is still numeric (int or float) and matches expected
            val_ok = isinstance(cell_val, (int, float)) and int(cell_val) == expected_val

            if fmt_ok and val_ok:
                print(f"PASS: Component {i} — {coord} has format '{cell_fmt}' and numeric value {cell_val} ({weight} pts)")
                total_score += weight
            elif fmt_ok and not val_ok:
                print(f"PARTIAL FAIL: Component {i} — {coord} has correct format '{cell_fmt}' but value is {cell_val!r} (type {type(cell_val).__name__}), expected numeric {expected_val}")
            elif not fmt_ok and val_ok:
                print(f"FAIL: Component {i} — {coord} has format '{cell_fmt}', expected '{EXPECTED_FORMAT}'")
            else:
                print(f"FAIL: Component {i} — {coord} format='{cell_fmt}' (expected '{EXPECTED_FORMAT}'), value={cell_val!r} (expected numeric {expected_val})")
        except Exception as e:
            print(f"ERROR: Component {i} — {coord}: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
