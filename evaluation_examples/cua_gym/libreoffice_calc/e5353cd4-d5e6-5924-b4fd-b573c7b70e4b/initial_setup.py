"""
Initial Setup: Bond Pricing and Sensitivity Table
Task ID: calc_fin_bond_pricing_046
Domain: libreoffice_calc

Creates an initial spreadsheet with bond parameters and an empty sensitivity table.
The agent must fill in the bond price formula and the sensitivity table formulas.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_bond_pricing_046'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: BondPricing ---
    ws = wb.active
    ws.title = 'BondPricing'

    # Section: Bond Parameters
    ws['A1'] = 'Face Value'
    ws['B1'] = 1000
    ws['A2'] = 'Coupon Rate'
    ws['B2'] = 0.05
    ws['A3'] = 'Years to Maturity'
    ws['B3'] = 10
    ws['A4'] = 'Required Yield'
    ws['B4'] = 0.06
    ws['A5'] = 'Bond Price'
    ws['B5'] = None  # Empty — agent must fill with =PV(B4,B3,-B1*B2,-B1)

    # Row 6: empty spacer

    # Row 7: Sensitivity Table section header
    ws['A7'] = 'Sensitivity Analysis: Bond Price by Yield and Maturity'

    # Row 8: Sensitivity table column headers (maturity values)
    ws['A8'] = 'Yield \\ Maturity'
    ws['B8'] = 5
    ws['C8'] = 7
    ws['D8'] = 10
    ws['E8'] = 15
    ws['F8'] = 20

    # Rows 9-13: Yield values in column A (inputs), empty prices in B-F
    yield_values = [0.04, 0.05, 0.06, 0.07, 0.08]
    for i, yld in enumerate(yield_values):
        row = 9 + i
        ws.cell(row=row, column=1, value=yld)
        # Columns B-F are intentionally empty — agent must fill sensitivity table

    # Format B2 and B4 as percentage for readability
    ws['B2'].number_format = '0.00%'
    ws['B4'].number_format = '0.00%'

    # Format B1 as currency
    ws['B1'].number_format = '$#,##0.00'

    # Format yield values in column A (A9:A13) as percentage
    for row in range(9, 14):
        ws.cell(row=row, column=1).number_format = '0.00%'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Structure:')
    print('  Sheet: BondPricing')
    print('  B1: 1000 (face value)')
    print('  B2: 0.05 (coupon rate)')
    print('  B3: 10 (years to maturity)')
    print('  B4: 0.06 (required yield)')
    print('  B5: empty (bond price — to be calculated)')
    print('  B8:F8: maturity headers (5, 7, 10, 15, 20)')
    print('  A9:A13: yield values (0.04 to 0.08)')
    print('  B9:F13: empty (sensitivity table — to be filled)')


create_initial()
