"""
Reward Script: Corporate Bond Pricing with Sensitivity Table
Task ID: calc_fin_bond_pricing_046
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: B5 has PV bond pricing formula with currency format         (0.25 pts)
  Component 2: Sensitivity table B9:F13 has correct PV formulas (all 25)  (0.35 pts)
  Component 3: Conditional formatting on B9:F13 (>1000 green, <1000 red) (0.20 pts)
  Component 4: Bold formatting on row 8 headers and column A yields       (0.10 pts)
  Component 5: Table borders (thin) on A8:F13                             (0.10 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_bond_pricing_046'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the sheet 'BondPricing' exists
    if 'BondPricing' not in wb.sheetnames:
        print("CRITICAL: Sheet 'BondPricing' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['BondPricing']

    # -------------------------------------------------------------------------
    # Component 1: B5 contains PV bond pricing formula with currency format
    # (0.25 points)
    # Expected: =PV(B4,B3,-B1*B2,-B1) in B5, formatted as currency ($#,##0.00)
    # This FAILS on initial (B5 is empty) -> PASSES on golden
    # -------------------------------------------------------------------------
    try:
        b5_val = ws['B5'].value
        b5_fmt = ws['B5'].number_format

        # Check that B5 has a PV formula (case-insensitive)
        has_pv_formula = (
            isinstance(b5_val, str) and
            b5_val.upper().replace(' ', '').startswith('=PV(')
        )

        # The formula should reference B4 (yield), B3 (maturity), B1 (face), B2 (coupon)
        formula_correct = False
        if has_pv_formula:
            normalized = b5_val.upper().replace(' ', '')
            # Must reference B4 (required yield), B3 (maturity), B1 (face value), B2 (coupon rate)
            formula_correct = (
                'B4' in normalized and
                'B3' in normalized and
                'B1' in normalized and
                'B2' in normalized
            )

        has_currency_fmt = (
            b5_fmt and (
                '$' in b5_fmt or
                '#,##0' in b5_fmt
            )
        )

        if formula_correct and has_currency_fmt:
            print(f"PASS: Component 1 — B5 has PV formula {repr(b5_val)} with currency format {repr(b5_fmt)} (0.25 pts)")
            total_score += 0.25
        elif formula_correct:
            # Partial: formula is correct but no currency format
            print(f"PASS (partial): Component 1 — B5 has correct PV formula but format is {repr(b5_fmt)}, expected currency (0.15 pts)")
            total_score += 0.15
        elif has_pv_formula:
            print(f"FAIL: Component 1 — B5 has PV formula but wrong references: {repr(b5_val)}")
        else:
            print(f"FAIL: Component 1 — B5 expected PV formula, found: {repr(b5_val)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Sensitivity table B9:F13 has correct PV formulas
    # (0.35 points)
    # Expected: each cell uses PV(row_yield, col_maturity, -$B$1*$B$2, -$B$1)
    # where row_yield = A9..A13 (relative to row), col_maturity = B8..F8 (relative to col)
    # Face value and coupon are absolute ($B$1, $B$2)
    # This FAILS on initial (B9:F13 are empty) -> PASSES on golden
    # -------------------------------------------------------------------------
    try:
        correct_count = 0
        total_cells = 25  # 5 rows x 5 cols

        for row in range(9, 14):
            for col in range(2, 7):
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(col)
                cell = ws.cell(row=row, column=col)
                val = cell.value

                if not isinstance(val, str):
                    print(f"  FAIL: {col_letter}{row} expected PV formula, found: {repr(val)}")
                    continue

                normalized = val.upper().replace(' ', '')

                if not normalized.startswith('=PV('):
                    print(f"  FAIL: {col_letter}{row} not a PV formula: {repr(val)}")
                    continue

                # Must reference row's yield column A (e.g., A9) relative — same row
                row_yield_ref = f'A{row}'
                # Must reference header maturity (same col, row 8) (e.g., B8)
                col_maturity_ref = f'{col_letter}8'
                # Must have absolute refs for face value and coupon
                has_abs_face = '$B$1' in val.upper() or '$B$1' in val
                has_abs_coupon = '$B$2' in val.upper() or '$B$2' in val
                has_row_yield = row_yield_ref.upper() in normalized
                has_col_maturity = col_maturity_ref.upper() in normalized

                if has_row_yield and has_col_maturity and has_abs_face and has_abs_coupon:
                    correct_count += 1
                else:
                    issues = []
                    if not has_row_yield:
                        issues.append(f'missing {row_yield_ref}')
                    if not has_col_maturity:
                        issues.append(f'missing {col_maturity_ref}')
                    if not has_abs_face:
                        issues.append('missing $B$1 (absolute face value)')
                    if not has_abs_coupon:
                        issues.append('missing $B$2 (absolute coupon rate)')
                    print(f"  FAIL: {col_letter}{row} formula issues: {'; '.join(issues)} — formula: {repr(val)}")

        # Award proportional score (all-or-nothing on 0.35 for full table, partial for partial)
        if correct_count == total_cells:
            print(f"PASS: Component 2 — All {total_cells} sensitivity table PV formulas correct (0.35 pts)")
            total_score += 0.35
        elif correct_count >= 20:
            print(f"PASS (partial): Component 2 — {correct_count}/{total_cells} PV formulas correct (0.25 pts)")
            total_score += 0.25
        elif correct_count >= 10:
            print(f"PASS (partial): Component 2 — {correct_count}/{total_cells} PV formulas correct (0.15 pts)")
            total_score += 0.15
        elif correct_count > 0:
            print(f"PASS (partial): Component 2 — {correct_count}/{total_cells} PV formulas correct (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 2 — No correct PV formulas found in sensitivity table B9:F13")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Conditional formatting on B9:F13
    # (0.20 points)
    # Expected: cells > 1000 get green fill, cells < 1000 get red fill
    # This FAILS on initial (no CF defined) -> PASSES on golden
    # -------------------------------------------------------------------------
    try:
        cf_rules = ws.conditional_formatting
        cf_list = list(cf_rules)

        # Find CF that applies to B9:F13 area
        found_greater_than = False
        found_less_than = False

        for cf in cf_list:
            cf_str = str(cf)
            # Check if this CF covers the sensitivity table area
            if 'B9' not in cf_str and 'B9:F13' not in cf_str:
                continue

            for rule in cf.rules:
                if rule.type == 'cellIs':
                    operator = getattr(rule, 'operator', None)
                    formula = getattr(rule, 'formula', [])
                    formula_val = formula[0] if formula else None

                    if operator == 'greaterThan' and formula_val == '1000':
                        # Check green fill
                        if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                            fg = rule.dxf.fill.fgColor
                            if fg:
                                # Accept any greenish fill (common greens: 92D050, 00FF00, etc.)
                                rgb_val = fg.rgb if fg.rgb else ''
                                is_green = (
                                    rgb_val.upper() in ['FF92D050', 'FF00FF00', '0092D050', '0000FF00'] or
                                    (len(rgb_val) >= 6 and
                                     int(rgb_val[-6:-4], 16) < 100 and  # low R
                                     int(rgb_val[-4:-2], 16) > 150)     # high G
                                )
                                if is_green:
                                    found_greater_than = True
                                    print(f"  PASS: CF greaterThan 1000 with green fill (rgb={rgb_val})")
                                else:
                                    # Accept any non-red fill for >1000 (could be light green)
                                    found_greater_than = True
                                    print(f"  PASS (lenient): CF greaterThan 1000 fill found (rgb={rgb_val})")
                            else:
                                found_greater_than = True
                                print(f"  PASS (no color check): CF greaterThan 1000 found")

                    elif operator == 'lessThan' and formula_val == '1000':
                        # Check red fill
                        if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                            fg = rule.dxf.fill.fgColor
                            if fg:
                                rgb_val = fg.rgb if fg.rgb else ''
                                is_red = (
                                    rgb_val.upper() in ['FFFF0000', '00FF0000'] or
                                    (len(rgb_val) >= 6 and
                                     int(rgb_val[-6:-4], 16) > 150 and  # high R
                                     int(rgb_val[-4:-2], 16) < 100)     # low G
                                )
                                if is_red:
                                    found_less_than = True
                                    print(f"  PASS: CF lessThan 1000 with red fill (rgb={rgb_val})")
                                else:
                                    found_less_than = True
                                    print(f"  PASS (lenient): CF lessThan 1000 fill found (rgb={rgb_val})")
                            else:
                                found_less_than = True
                                print(f"  PASS (no color check): CF lessThan 1000 found")

        if found_greater_than and found_less_than:
            print(f"PASS: Component 3 — Conditional formatting on B9:F13 (>1000 green, <1000 red) (0.20 pts)")
            total_score += 0.20
        elif found_greater_than or found_less_than:
            print(f"PASS (partial): Component 3 — Only one CF rule found (greaterThan={found_greater_than}, lessThan={found_less_than}) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — No conditional formatting found on B9:F13 (initial has no CF)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: Bold formatting on row 8 (headers) and column A (yield values)
    # (0.10 points)
    # Row 8 (A8:F8) and column A rows 9:13 should be bold
    # This FAILS on initial (no bold) -> PASSES on golden
    # -------------------------------------------------------------------------
    try:
        # Check row 8 bold: A8 to F8
        row8_bold_count = 0
        for col in range(1, 7):
            cell = ws.cell(row=8, column=col)
            if cell.font.bold:
                row8_bold_count += 1

        # Check column A bold: A9 to A13
        colA_bold_count = 0
        for row in range(9, 14):
            cell = ws.cell(row=row, column=1)
            if cell.font.bold:
                colA_bold_count += 1

        total_bold = row8_bold_count + colA_bold_count
        expected_bold = 6 + 5  # 6 header cells + 5 yield cells

        if row8_bold_count == 6 and colA_bold_count == 5:
            print(f"PASS: Component 4 — Row 8 ({row8_bold_count}/6 bold) and col A ({colA_bold_count}/5 bold) (0.10 pts)")
            total_score += 0.10
        elif total_bold >= 8:
            print(f"PASS (partial): Component 4 — {total_bold}/{expected_bold} bold cells (row8={row8_bold_count}/6, colA={colA_bold_count}/5) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4 — Insufficient bold formatting: row8={row8_bold_count}/6, colA={colA_bold_count}/5 (initial has 0 bold)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -------------------------------------------------------------------------
    # Component 5: Table borders on A8:F13 (thin borders on all cells)
    # (0.10 points)
    # This FAILS on initial (no borders) -> PASSES on golden
    # -------------------------------------------------------------------------
    try:
        border_count = 0
        total_border_cells = 6 * 6  # 6 rows (8-13) x 6 cols (A-F) = 36

        for row in range(8, 14):
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                b = cell.border
                # Check at least one border side is set
                has_any_border = any([
                    b.left.style is not None,
                    b.right.style is not None,
                    b.top.style is not None,
                    b.bottom.style is not None
                ])
                if has_any_border:
                    border_count += 1

        if border_count == total_border_cells:
            print(f"PASS: Component 5 — All {total_border_cells} cells in A8:F13 have borders (0.10 pts)")
            total_score += 0.10
        elif border_count >= 30:
            print(f"PASS (partial): Component 5 — {border_count}/{total_border_cells} cells have borders (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — Only {border_count}/{total_border_cells} cells have borders (initial has 0 borders)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
