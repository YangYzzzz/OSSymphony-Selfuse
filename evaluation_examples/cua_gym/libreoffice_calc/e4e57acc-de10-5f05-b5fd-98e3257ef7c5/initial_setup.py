"""
Initial Setup: Set print area and repeating header rows for Invoice sheet
Task ID: calc_gg1_017
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_017'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Invoice'

    # --- Row 1: Company logo placeholder (merged across columns) ---
    ws.merge_cells('A1:G1')
    ws['A1'] = 'MERIDIAN GLOBAL SUPPLIES INC.'
    ws['A1'].font = Font(name='Arial', size=18, bold=True, color='1F4E79')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36

    # --- Row 2: Column headers ---
    headers = ['Item #', 'Description', 'Qty', 'Unit Price', 'Discount %', 'Tax %', 'Line Total']
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[2].height = 28

    # --- Column widths ---
    ws.column_dimensions['A'].width = 10   # Item #
    ws.column_dimensions['B'].width = 40   # Description
    ws.column_dimensions['C'].width = 8    # Qty
    ws.column_dimensions['D'].width = 14   # Unit Price
    ws.column_dimensions['E'].width = 12   # Discount %
    ws.column_dimensions['F'].width = 10   # Tax %
    ws.column_dimensions['G'].width = 16   # Line Total

    # --- Rows 3-45: Line items (43 items) ---
    descriptions = [
        'Premium A4 Copy Paper (5-ream box)',
        'Heavy-Duty Stapler - Industrial Grade',
        'Ergonomic Office Chair - Mesh Back',
        'LED Desk Lamp - Adjustable Arm',
        'Wireless Keyboard & Mouse Combo',
        'Whiteboard Markers Assorted (12-pack)',
        'Manila File Folders - Letter Size (100-pack)',
        'Ink Cartridge - Black (High Yield)',
        'Ink Cartridge - Cyan (Standard)',
        'Ink Cartridge - Magenta (Standard)',
        'Ink Cartridge - Yellow (Standard)',
        'USB-C Docking Station - Dual Monitor',
        'Noise-Cancelling Headset with Mic',
        'Standing Desk Converter - 36 inch',
        'Cable Management Kit - Under Desk',
        'Surge Protector Power Strip (8-outlet)',
        'External SSD 1TB - USB 3.2',
        'Webcam 1080p HD - Auto Focus',
        'Document Scanner - Duplex Feed',
        'Laminating Machine - A3 Compatible',
        'Laminating Pouches A4 (200-pack)',
        'Paper Shredder - Cross Cut',
        'Desk Organizer - 5 Compartment',
        'Sticky Notes Assorted Colors (24-pack)',
        'Ballpoint Pens - Blue (50-pack)',
        'Gel Pens - Black (24-pack)',
        'Binder Clips Assorted Sizes (60-pack)',
        'Presentation Clicker - Laser Pointer',
        'HDMI Cable 2m - 4K Support',
        'USB Flash Drive 64GB (5-pack)',
        'Monitor Stand Riser - Bamboo',
        'Footrest - Adjustable Angle',
        'Desk Calendar 2025 - Monthly Planner',
        'Push Pins - Assorted Colors (200-pack)',
        'Rubber Bands Assorted (500g bag)',
        'Correction Tape (10-pack)',
        'Highlighters Assorted (12-pack)',
        'Scissors - Stainless Steel 8 inch',
        'Tape Dispenser with 3 Rolls',
        'Envelope - C5 Self-Seal (500-pack)',
        'Printer Paper A3 (500 sheets)',
        'Toner Cartridge - Black (Compatible)',
        'Air Duster Compressed (6-pack)',
    ]

    quantities = [
        20, 5, 3, 8, 10, 15, 6, 4, 4, 4,
        4, 2, 6, 1, 5, 7, 3, 4, 1, 1,
        10, 2, 8, 20, 12, 8, 10, 3, 6, 5,
        4, 3, 10, 8, 5, 15, 12, 6, 5, 3,
        8, 2, 10,
    ]

    unit_prices = [
        34.99, 28.50, 349.00, 45.75, 39.99, 12.50, 22.00, 42.99, 28.99, 28.99,
        28.99, 189.00, 75.50, 429.00, 18.95, 24.99, 109.99, 64.50, 299.00, 179.00,
        15.99, 149.99, 14.50, 8.99, 16.75, 11.99, 7.50, 34.99, 9.99, 24.99,
        39.99, 44.50, 6.99, 4.50, 3.99, 5.99, 9.50, 8.99, 6.50, 45.00,
        32.50, 89.99, 12.99,
    ]

    discounts = [
        5, 0, 10, 0, 5, 0, 0, 0, 0, 0,
        0, 8, 5, 12, 0, 0, 5, 0, 10, 0,
        0, 5, 0, 10, 0, 0, 0, 0, 0, 5,
        0, 0, 0, 0, 0, 0, 5, 0, 0, 0,
        0, 3, 0,
    ]

    tax_rates = [
        10, 10, 10, 10, 10, 10, 10, 10, 10, 10,
        10, 10, 10, 10, 10, 10, 10, 10, 10, 10,
        10, 10, 10, 10, 10, 10, 10, 10, 10, 10,
        10, 10, 10, 10, 10, 10, 10, 10, 10, 10,
        10, 10, 10,
    ]

    data_font = Font(name='Arial', size=10)
    currency_format = '$#,##0.00'
    pct_format = '0%'
    number_format_int = '0'

    for i in range(43):
        row = i + 3
        item_num = f'INV-{1001 + i}'
        qty = quantities[i]
        price = unit_prices[i]
        disc = discounts[i]
        tax = tax_rates[i]
        # Calculate line total: qty * price * (1 - disc/100) * (1 + tax/100)
        line_total = round(qty * price * (1 - disc / 100) * (1 + tax / 100), 2)

        ws.cell(row=row, column=1, value=item_num).font = data_font
        ws.cell(row=row, column=2, value=descriptions[i]).font = data_font
        cell_qty = ws.cell(row=row, column=3, value=qty)
        cell_qty.font = data_font
        cell_qty.alignment = Alignment(horizontal='center')
        cell_qty.number_format = number_format_int

        cell_price = ws.cell(row=row, column=4, value=price)
        cell_price.font = data_font
        cell_price.number_format = currency_format

        cell_disc = ws.cell(row=row, column=5, value=disc / 100)
        cell_disc.font = data_font
        cell_disc.number_format = pct_format
        cell_disc.alignment = Alignment(horizontal='center')

        cell_tax = ws.cell(row=row, column=6, value=tax / 100)
        cell_tax.font = data_font
        cell_tax.number_format = pct_format
        cell_tax.alignment = Alignment(horizontal='center')

        cell_total = ws.cell(row=row, column=7, value=line_total)
        cell_total.font = data_font
        cell_total.number_format = currency_format

        # Light alternating row background
        if i % 2 == 1:
            alt_fill = PatternFill(start_color='FFD9E2F3', end_color='FFD9E2F3', fill_type='solid')
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = alt_fill

    # NO print area set - that's the task
    # NO repeating rows set - that's the task

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
