"""
Initial Setup: Add a thin right border to column A cells (A1:A20)
Task ID: calc_fmt_border_right_only_074
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_border_right_only_074'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Comparison Table'

    # Row labels for column A (A1:A20)
    row_labels = [
        'Category',
        'Option A',
        'Option B',
        'Option C',
        'Option D',
        'Option E',
        'Option F',
        'Option G',
        'Option H',
        'Option I',
        'Option J',
        'Option K',
        'Option L',
        'Option M',
        'Option N',
        'Option O',
        'Option P',
        'Option Q',
        'Option R',
        'Option S',
    ]

    # Column headers for B-F
    col_headers = ['Cost ($)', 'Quality', 'Reliability', 'Support', 'Rating']

    # Set column A header
    # Row labels already set below

    # Comparison data (18 data rows after header row)
    comparison_data = [
        # Cost,  Quality, Reliability, Support, Rating
        [1200,   8.5,     9.0,         7.5,     8.2],
        [950,    7.0,     8.5,         8.0,     7.8],
        [1450,   9.2,     8.8,         9.0,     9.0],
        [780,    6.5,     7.0,         6.5,     6.8],
        [1100,   8.0,     8.0,         8.5,     8.1],
        [1350,   8.8,     9.2,         8.0,     8.7],
        [670,    6.0,     6.5,         7.0,     6.5],
        [1600,   9.5,     9.5,         9.5,     9.5],
        [890,    7.5,     7.8,         7.0,     7.4],
        [1020,   7.8,     8.2,         8.0,     7.9],
        [1250,   8.3,     8.6,         8.3,     8.4],
        [740,    6.2,     6.8,         6.0,     6.3],
        [1480,   9.0,     9.1,         8.8,     9.0],
        [830,    7.2,     7.5,         7.2,     7.3],
        [1150,   8.1,     8.4,         7.8,     8.1],
        [975,    7.6,     8.0,         7.5,     7.7],
        [1380,   8.9,     9.0,         8.7,     8.9],
        [700,    6.3,     6.7,         6.2,     6.4],
        [1080,   7.9,     8.1,         7.8,     7.9],
    ]

    # Write header row (A1 + B1:F1)
    ws['A1'] = 'Category'
    for col_idx, header in enumerate(col_headers, 2):
        ws.cell(row=1, column=col_idx, value=header)

    # Style header row - bold font
    header_font = Font(bold=True)
    for col_idx in range(1, 7):
        ws.cell(row=1, column=col_idx).font = header_font

    # Write row labels A2:A20 and data B2:F20
    for row_idx, (label, data_row) in enumerate(zip(row_labels[1:], comparison_data), 2):
        ws.cell(row=row_idx, column=1, value=label)
        for col_idx, val in enumerate(data_row, 2):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 9

    # NO borders on any cells (this is the initial state)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Comparison Table')
    print(f'  Rows: 20 (1 header + 19 data rows)')
    print(f'  Columns: A-F (6 columns)')
    print(f'  Borders: NONE (task requires adding right border to A1:A20)')


create_initial()
