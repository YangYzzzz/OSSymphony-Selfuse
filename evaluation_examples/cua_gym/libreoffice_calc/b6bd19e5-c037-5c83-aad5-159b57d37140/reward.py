"""
Reward Script: Add a thin right border to column A cells (A1:A20) to create
               a visible separator between row labels and data columns.
Task ID: calc_fmt_border_right_only_074
Domain: libreoffice_calc
Scoring:
  Component 1: At least half of A1:A20 cells have thin right border        (0.5 pts)
  Component 2: ALL A1:A20 have thin right border AND no left/top/bottom    (0.3 pts)
  Component 3: No unintended borders on non-A-column cells (B1:F20)        (0.2 pts)
               — awarded only when Component 1 passes (conditional)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmt_border_right_only_074'
SHEET_NAME = 'Comparison Table'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet 'Comparison Table' must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found in workbook. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Component 1: At least 10/20 cells in A1:A20 have a thin right border (0.5 points)
    # This FAILS on initial (0 borders) and earns partial/full credit on golden
    try:
        cells_with_right_border = 0
        missing_cells = []
        for row in range(1, 21):
            cell = ws.cell(row=row, column=1)
            if cell.border.right.style == 'thin':
                cells_with_right_border += 1
            else:
                missing_cells.append(f"A{row}")

        if cells_with_right_border == 20:
            print(f"PASS: Component 1 — All 20 cells A1:A20 have thin right border (0.5 pts)")
            total_score += 0.5
        elif cells_with_right_border >= 10:
            # Partial credit for partial completion
            partial = round(0.5 * (cells_with_right_border / 20), 2)
            print(f"PARTIAL: Component 1 — {cells_with_right_border}/20 cells have thin right border. "
                  f"Missing: {missing_cells[:5]}{'...' if len(missing_cells) > 5 else ''} "
                  f"({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {cells_with_right_border}/20 cells in A1:A20 have a thin right "
                  f"border. Expected thin right border style on all 20 cells.")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        cells_with_right_border = 0

    # Component 2: ALL A1:A20 have thin right border AND no left/top/bottom borders (0.3 points)
    # This compound condition: right=thin AND left/top/bottom=None for all 20 cells
    # FAILS on initial: no right borders → compound always fails (cells_right_only_count < 20)
    # PASSES on golden: right=thin and left/top/bottom=None for all 20 cells
    try:
        cells_right_only_count = 0
        failure_details = []
        for row in range(1, 21):
            cell = ws.cell(row=row, column=1)
            b = cell.border
            has_right = (b.right.style == 'thin')
            no_other = (b.left.style is None and b.top.style is None and b.bottom.style is None)
            if has_right and no_other:
                cells_right_only_count += 1
            else:
                failure_details.append(
                    f"A{row}(right={b.right.style}, left={b.left.style}, "
                    f"top={b.top.style}, bottom={b.bottom.style})"
                )

        if cells_right_only_count == 20:
            print(f"PASS: Component 2 — All A1:A20 have ONLY right border (thin), no other borders (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — Only {cells_right_only_count}/20 cells have right-only border pattern. "
                  f"Issues: {failure_details[:3]}{'...' if len(failure_details) > 3 else ''}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: No borders added to other cells (B1:F20 range) (0.2 points)
    # Awarded only when at least one A-column cell has the right border
    # This ensures it fails on the initial file (no right borders → not conditioned on task success)
    # To make this fail on initial, we gate it: only check if right-border cells > 0
    # But instead, we check the CONJUNCTION: (all A1:A20 have right border) AND (B1:F20 clean)
    # This FAILS on initial because A1:A20 don't have right borders
    try:
        if cells_with_right_border == 0:
            print(f"FAIL: Component 3 — Skipped (no right borders set on A1:A20 — task not started)")
        else:
            other_cells_with_borders = []
            for row in range(1, 21):
                for col in range(2, 7):  # columns B through F
                    cell = ws.cell(row=row, column=col)
                    b = cell.border
                    if any([b.left.style, b.right.style, b.top.style, b.bottom.style]):
                        from openpyxl.utils import get_column_letter
                        coord = f"{get_column_letter(col)}{row}"
                        other_cells_with_borders.append(coord)

            if not other_cells_with_borders:
                print(f"PASS: Component 3 — No unintended borders on B1:F20 cells (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 3 — Unexpected borders found on non-column-A cells: "
                      f"{other_cells_with_borders[:5]}{'...' if len(other_cells_with_borders) > 5 else ''}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
