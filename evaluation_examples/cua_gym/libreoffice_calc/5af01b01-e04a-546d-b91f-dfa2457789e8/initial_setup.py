"""
Initial Setup: Configure print setup on Annual Report sheet
Task ID: calc_ggf_030
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_030'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Annual Report'

    # --- Header Styling ---
    header_font = Font(name='Arial', size=14, bold=True, color='1F4E79')
    subheader_font = Font(name='Arial', size=11, bold=True, color='2E75B6')
    col_header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    col_header_fill = PatternFill(start_color='FF2E75B6', end_color='FF2E75B6', fill_type='solid')
    col_header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_align = Alignment(horizontal='right', vertical='center')
    text_align = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='B0B0B0'),
        right=Side(style='thin', color='B0B0B0'),
        top=Side(style='thin', color='B0B0B0'),
        bottom=Side(style='thin', color='B0B0B0'),
    )

    # Row 1: Title (merged A1:N1)
    ws.merge_cells('A1:N1')
    ws['A1'] = 'Pinnacle Industries Inc. — Annual Financial Report 2024'
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Row 2: Subtitle (merged A2:N2)
    ws.merge_cells('A2:N2')
    ws['A2'] = 'Departmental Budget Performance & Variance Analysis (All figures in USD thousands)'
    ws['A2'].font = subheader_font
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22

    # Row 3: Column headers
    headers = [
        'Department', 'Category',
        'Q1 Budget', 'Q1 Actual', 'Q1 Variance',
        'Q2 Budget', 'Q2 Actual', 'Q2 Variance',
        'Q3 Budget', 'Q3 Actual', 'Q3 Variance',
        'Q4 Budget', 'Q4 Actual', 'Q4 Variance',
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = col_header_align
        cell.border = thin_border
    ws.row_dimensions[3].height = 28

    # Column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 20
    for col_letter in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws.column_dimensions[col_letter].width = 13

    # --- Data rows (rows 4-60) ---
    departments = [
        'Engineering', 'Marketing', 'Sales', 'Finance',
        'Human Resources', 'Operations', 'Research & Dev',
        'Customer Support', 'Legal', 'IT Infrastructure',
        'Product Management', 'Quality Assurance',
    ]

    categories = [
        'Personnel', 'Equipment', 'Travel', 'Training',
        'Software Licenses',
    ]

    # Alternating row fill
    alt_fill_1 = PatternFill(start_color='FFF2F7FB', end_color='FFF2F7FB', fill_type='solid')
    alt_fill_2 = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')

    import random
    random.seed(42)

    row_num = 4
    for dept_idx, dept in enumerate(departments):
        for cat_idx, cat in enumerate(categories):
            if row_num > 60:
                break

            cell_a = ws.cell(row=row_num, column=1, value=dept)
            cell_a.alignment = text_align
            cell_a.border = thin_border

            cell_b = ws.cell(row=row_num, column=2, value=cat)
            cell_b.alignment = text_align
            cell_b.border = thin_border

            # Generate quarterly budget/actual/variance data
            for q in range(4):
                base_budget = random.randint(80, 500)
                variance_pct = random.uniform(-0.15, 0.12)
                actual = round(base_budget * (1 + variance_pct), 1)
                variance = round(actual - base_budget, 1)

                budget_col = 3 + q * 3
                actual_col = 4 + q * 3
                variance_col = 5 + q * 3

                cell_bud = ws.cell(row=row_num, column=budget_col, value=base_budget)
                cell_bud.number_format = '#,##0.0'
                cell_bud.alignment = data_align
                cell_bud.border = thin_border

                cell_act = ws.cell(row=row_num, column=actual_col, value=actual)
                cell_act.number_format = '#,##0.0'
                cell_act.alignment = data_align
                cell_act.border = thin_border

                cell_var = ws.cell(row=row_num, column=variance_col, value=variance)
                cell_var.number_format = '#,##0.0'
                cell_var.alignment = data_align
                cell_var.border = thin_border

                # Color negative variances red, positive green
                if variance < 0:
                    cell_var.font = Font(color='CC0000')
                else:
                    cell_var.font = Font(color='008000')

            # Alternating row fill
            fill = alt_fill_1 if (row_num % 2 == 0) else alt_fill_2
            for c in range(1, 15):
                ws.cell(row=row_num, column=c).fill = fill

            row_num += 1
        if row_num > 60:
            break

    # Fill remaining rows if needed (up to row 60)
    extra_depts = ['Procurement', 'Facilities', 'Public Relations']
    extra_idx = 0
    while row_num <= 60:
        dept = extra_depts[extra_idx % len(extra_depts)]
        cat = categories[(row_num - 4) % len(categories)]
        extra_idx += 1

        cell_a = ws.cell(row=row_num, column=1, value=dept)
        cell_a.alignment = text_align
        cell_a.border = thin_border

        cell_b = ws.cell(row=row_num, column=2, value=cat)
        cell_b.alignment = text_align
        cell_b.border = thin_border

        for q in range(4):
            base_budget = random.randint(60, 350)
            variance_pct = random.uniform(-0.12, 0.10)
            actual = round(base_budget * (1 + variance_pct), 1)
            variance = round(actual - base_budget, 1)

            budget_col = 3 + q * 3
            actual_col = 4 + q * 3
            variance_col = 5 + q * 3

            cell_bud = ws.cell(row=row_num, column=budget_col, value=base_budget)
            cell_bud.number_format = '#,##0.0'
            cell_bud.alignment = data_align
            cell_bud.border = thin_border

            cell_act = ws.cell(row=row_num, column=actual_col, value=actual)
            cell_act.number_format = '#,##0.0'
            cell_act.alignment = data_align
            cell_act.border = thin_border

            cell_var = ws.cell(row=row_num, column=variance_col, value=variance)
            cell_var.number_format = '#,##0.0'
            cell_var.alignment = data_align
            cell_var.border = thin_border

            if variance < 0:
                cell_var.font = Font(color='CC0000')
            else:
                cell_var.font = Font(color='008000')

        fill = alt_fill_1 if (row_num % 2 == 0) else alt_fill_2
        for c in range(1, 15):
            ws.cell(row=row_num, column=c).fill = fill

        row_num += 1

    # NO print setup configured — that is the task
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
