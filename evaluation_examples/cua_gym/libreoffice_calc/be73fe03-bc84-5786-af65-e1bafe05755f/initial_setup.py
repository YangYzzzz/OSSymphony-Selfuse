"""
Initial Setup: Create a multi-sheet workbook template for department heads (unprotected).
Task ID: calc_gsi_018
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_018'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # Styling helpers
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    def style_header(ws, headers, row=1):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    # --- Sheet 1: Revenue ---
    ws1 = wb.active
    ws1.title = "Revenue"
    rev_headers = ["Region", "Q1 2025", "Q2 2025", "Q3 2025", "Q4 2025", "Annual Total"]
    style_header(ws1, rev_headers)
    rev_data = [
        ["North America", 1250000, 1380000, 1425000, 1510000],
        ["Europe", 890000, 920000, 975000, 1020000],
        ["Asia Pacific", 675000, 710000, 780000, 845000],
        ["Latin America", 320000, 345000, 360000, 395000],
        ["Middle East & Africa", 185000, 195000, 210000, 230000],
        ["Central Europe", 445000, 462000, 490000, 515000],
        ["Southeast Asia", 290000, 315000, 340000, 375000],
        ["Scandinavia", 180000, 192000, 205000, 218000],
        ["Oceania", 155000, 168000, 178000, 192000],
        ["South Asia", 210000, 228000, 245000, 268000],
        ["Eastern Europe", 125000, 138000, 150000, 165000],
        ["Caribbean", 88000, 95000, 102000, 112000],
    ]
    for r, row_data in enumerate(rev_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            if c >= 2:
                cell.number_format = '$#,##0'
        # Annual Total formula
        ws1.cell(row=r, column=6, value=f'=SUM(B{r}:E{r})')
        ws1.cell(row=r, column=6).number_format = '$#,##0'

    ws1.column_dimensions["A"].width = 22
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws1.column_dimensions[col_letter].width = 15

    # --- Sheet 2: Expenses ---
    ws2 = wb.create_sheet("Expenses")
    exp_headers = ["Department", "Salaries", "Operations", "Marketing", "Travel", "Total"]
    style_header(ws2, exp_headers)
    exp_data = [
        ["Engineering", 2450000, 380000, 45000, 120000],
        ["Sales", 1850000, 210000, 520000, 340000],
        ["Marketing", 980000, 165000, 890000, 185000],
        ["Human Resources", 620000, 95000, 25000, 55000],
        ["Finance", 780000, 120000, 15000, 42000],
        ["Legal", 540000, 85000, 10000, 65000],
        ["Product", 1320000, 245000, 75000, 98000],
        ["Customer Support", 890000, 175000, 35000, 28000],
        ["Research", 1150000, 420000, 30000, 145000],
        ["Operations", 720000, 310000, 20000, 38000],
        ["IT Infrastructure", 680000, 520000, 12000, 32000],
    ]
    for r, row_data in enumerate(exp_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if c >= 2:
                cell.number_format = '$#,##0'
        ws2.cell(row=r, column=6, value=f'=SUM(B{r}:E{r})')
        ws2.cell(row=r, column=6).number_format = '$#,##0'

    ws2.column_dimensions["A"].width = 20
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws2.column_dimensions[col_letter].width = 15

    # --- Sheet 3: Headcount ---
    ws3 = wb.create_sheet("Headcount")
    hc_headers = ["Department", "Full-Time", "Part-Time", "Contractors", "Total", "Budget Utilization"]
    style_header(ws3, hc_headers)
    hc_data = [
        ["Engineering", 85, 12, 25, None, "92%"],
        ["Sales", 62, 8, 15, None, "88%"],
        ["Marketing", 34, 6, 18, None, "95%"],
        ["Human Resources", 18, 3, 4, None, "76%"],
        ["Finance", 22, 2, 6, None, "82%"],
        ["Legal", 14, 1, 8, None, "90%"],
        ["Product", 45, 5, 12, None, "87%"],
        ["Customer Support", 38, 15, 20, None, "93%"],
        ["Research", 28, 4, 10, None, "85%"],
        ["Operations", 20, 6, 8, None, "79%"],
    ]
    for r, row_data in enumerate(hc_data, 2):
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r, column=c, value=val)
        # Total formula
        ws3.cell(row=r, column=5, value=f'=SUM(B{r}:D{r})')

    ws3.column_dimensions["A"].width = 20
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws3.column_dimensions[col_letter].width = 18

    # --- Sheet 4: KPIs ---
    ws4 = wb.create_sheet("KPIs")
    kpi_headers = ["KPI Metric", "Target", "Actual", "Variance", "Status"]
    style_header(ws4, kpi_headers)
    kpi_data = [
        ["Revenue Growth (%)", 15.0, 12.8, -2.2, "Below Target"],
        ["Customer Acquisition Cost ($)", 250, 235, 15, "On Track"],
        ["Customer Retention Rate (%)", 92.0, 94.5, 2.5, "Exceeds"],
        ["Net Promoter Score", 45, 52, 7, "Exceeds"],
        ["Employee Satisfaction", 4.2, 4.0, -0.2, "Below Target"],
        ["Operating Margin (%)", 18.0, 17.2, -0.8, "Below Target"],
        ["Time to Market (days)", 90, 78, 12, "Exceeds"],
        ["Defect Rate (%)", 2.0, 1.5, 0.5, "Exceeds"],
        ["Support Resolution (hrs)", 24, 18, 6, "Exceeds"],
        ["Market Share (%)", 12.0, 11.5, -0.5, "Below Target"],
        ["R&D Spend Ratio (%)", 20.0, 21.3, 1.3, "On Track"],
        ["Inventory Turnover", 8.0, 7.6, -0.4, "Below Target"],
    ]
    for r, row_data in enumerate(kpi_data, 2):
        for c, val in enumerate(row_data, 1):
            ws4.cell(row=r, column=c, value=val)

    ws4.column_dimensions["A"].width = 30
    for col_letter in ["B", "C", "D", "E"]:
        ws4.column_dimensions[col_letter].width = 16

    # Workbook is NOT protected (initial state)
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
