"""
Reward Script: Import CSV file into LibreOffice Calc spreadsheet
Task ID: calc_adv_import_csv_041
Domain: libreoffice_calc
Scoring:
  Component 1: Headers in row 1 (Date, Region, Product, Quantity, Price, Total) — 0.3 pts
  Component 2: Row count is 501 (header + 500 data rows) — 0.3 pts
  Component 3: Columns correctly separated (6 columns) and numeric columns contain numbers — 0.4 pts
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_adv_import_csv_041'

EXPECTED_HEADERS = ['Date', 'Region', 'Product', 'Quantity', 'Price', 'Total']
EXPECTED_ROWS = 501   # 1 header + 500 data rows
EXPECTED_COLS = 6

# Numeric columns: Quantity (col 4), Price (col 5), Total (col 6)
NUMERIC_COLS = [4, 5, 6]


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active
    if ws is None:
        print("CRITICAL: No active sheet found")
        print("REWARD: 0.0")
        return 0.0

    # Snapshot dimensions BEFORE any cell access (accessing cells beyond the used range
    # can expand max_row/max_column as an openpyxl side effect)
    initial_max_row = ws.max_row
    initial_max_col = ws.max_column

    # Component 1: Headers in row 1 are correct (0.3 points)
    # The initial file has no headers (empty). After CSV import, row 1 should have the 6 column headers.
    try:
        actual_headers = [ws.cell(row=1, column=col).value for col in range(1, EXPECTED_COLS + 1)]
        if actual_headers == EXPECTED_HEADERS:
            print(f"PASS: Component 1 — headers correct in row 1: {actual_headers} (0.3 pts)")
            total_score += 0.3
        else:
            # Check case-insensitive match
            if actual_headers and all(
                h is not None and str(h).strip().lower() == exp.lower()
                for h, exp in zip(actual_headers, EXPECTED_HEADERS)
            ) and len(actual_headers) == len(EXPECTED_HEADERS):
                print(f"PASS: Component 1 — headers correct (case-insensitive) in row 1: {actual_headers} (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 1 — expected headers {EXPECTED_HEADERS}, found {actual_headers}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Row count is 501 (header row + 500 data rows) (0.3 points)
    # The initial file has 1 empty row. After CSV import, should have 501 rows.
    # Use the snapshotted row count (before cell access side effects).
    try:
        actual_rows = initial_max_row
        if actual_rows == EXPECTED_ROWS:
            print(f"PASS: Component 2 — row count correct: {actual_rows} rows (header + 500 data rows) (0.3 pts)")
            total_score += 0.3
        elif actual_rows and actual_rows >= 490:
            # Very close to expected — give feedback without score
            print(f"FAIL: Component 2 — expected {EXPECTED_ROWS} rows, found {actual_rows} rows")
        else:
            print(f"FAIL: Component 2 — expected {EXPECTED_ROWS} rows, found {actual_rows} rows (data likely not imported)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Columns correctly separated (6 cols) and numeric columns contain numbers (0.4 points)
    # The initial file has 1 column with None. After import, should have 6 separate columns
    # with Quantity, Price, Total as numeric types (not text strings).
    # Use snapshotted column count to avoid openpyxl side-effect expansion.
    try:
        actual_cols = initial_max_col

        # Sub-check A: 6 columns present
        cols_ok = (actual_cols == EXPECTED_COLS)

        # Sub-check B: Numeric columns contain actual numbers (not text strings)
        # Check across multiple rows to ensure data is truly numeric, not just one row
        numeric_ok = True
        non_numeric_count = 0
        if cols_ok and initial_max_row and initial_max_row > 1:
            rows_to_check = min(initial_max_row, 11)  # check up to first 10 data rows
            for row in range(2, rows_to_check + 1):
                for col in NUMERIC_COLS:
                    val = ws.cell(row=row, column=col).value
                    if val is not None and not isinstance(val, (int, float)):
                        # If it's a string that cannot be parsed as a number, it's non-numeric
                        try:
                            float(str(val).replace(',', ''))
                        except (ValueError, TypeError):
                            non_numeric_count += 1
                    elif val is None:
                        non_numeric_count += 1

            # Allow zero non-numeric values
            numeric_ok = (non_numeric_count == 0)

        if cols_ok and numeric_ok:
            print(f"PASS: Component 3 — {actual_cols} columns detected, numeric columns contain actual numbers (0.4 pts)")
            total_score += 0.4
        elif cols_ok and not numeric_ok:
            print(f"FAIL: Component 3 — {actual_cols} columns OK but {non_numeric_count} non-numeric values found in Quantity/Price/Total columns")
        elif not cols_ok:
            print(f"FAIL: Component 3 — expected {EXPECTED_COLS} columns, found {actual_cols} (CSV may not be split by delimiter)")
        else:
            print(f"FAIL: Component 3 — cols={actual_cols}, non_numeric_count={non_numeric_count}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
