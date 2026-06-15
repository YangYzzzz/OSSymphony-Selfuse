"""
Reward Script: Create pivot table from enrollment data showing count of students by Department and Course Level
Task ID: calc_edu_enrollment_pivot_013
Domain: libreoffice_calc
Scoring:
  - Component 1: 'Enrollment Summary' sheet exists (0.3 pts)
  - Component 2: Correct column headers (Department, Level 100/200/300/400, Grand Total) (0.2 pts)
  - Component 3: Correct department data rows with accurate counts (0.3 pts)
  - Component 4: Grand total row with correct totals (0.2 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_enrollment_pivot_013'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: 'Enrollment Summary' sheet exists (0.3 points)
    # Initial file has NO 'Enrollment Summary' sheet; golden file has it.
    try:
        if 'Enrollment Summary' in wb.sheetnames:
            print("PASS: Component 1 — 'Enrollment Summary' sheet exists (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — 'Enrollment Summary' sheet not found. Sheets: {wb.sheetnames}")
            # Cannot check further components if sheet doesn't exist
            final_score = min(total_score, 1.0)
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {final_score}")
            return final_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    ws = wb['Enrollment Summary']

    # Component 2: Correct column headers (0.2 points)
    # Expected: A row contains Department in col 1 AND level identifiers (100/200/300/400) in subsequent cols
    # The title row may contain "Department" as part of a longer phrase; we need a row where
    # col 1 is exactly/primarily "Department" and cols 2+ have level markers.
    try:
        header_row = None
        # Search for the header row within first 6 rows
        # A header row has 'department' in col 1 AND at least one of 100/200/300/400 in other cols
        for r in range(1, 7):
            c1_val = str(ws.cell(row=r, column=1).value or '').strip().lower()
            if 'department' not in c1_val:
                continue
            # Check if other columns in this row have level numbers
            row_headers = [str(ws.cell(row=r, column=c).value or '').lower() for c in range(2, 7)]
            row_text = ' '.join(row_headers)
            if '100' in row_text or '200' in row_text or '300' in row_text or '400' in row_text:
                header_row = r
                break

        if header_row is not None:
            # Check column headers in the header row
            headers = [ws.cell(row=header_row, column=c).value for c in range(1, 7)]
            headers_str = [str(h).strip().lower() if h is not None else '' for h in headers]

            # Check department in first column
            has_dept = 'department' in headers_str[0]

            # Check that level 100, 200, 300, 400 are represented in columns 2-5
            header_text = ' '.join(headers_str)
            has_100 = '100' in header_text
            has_200 = '200' in header_text
            has_300 = '300' in header_text
            has_400 = '400' in header_text

            if has_dept and has_100 and has_200 and has_300 and has_400:
                print(f"PASS: Component 2 — Column headers found: {headers} (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 2 — Unexpected column headers: {headers}")
                print(f"  has_dept={has_dept}, has_100={has_100}, has_200={has_200}, has_300={has_300}, has_400={has_400}")
        else:
            print(f"FAIL: Component 2 — No header row with 'Department' + level columns found in first 6 rows")
            # Print all cell values to help debug
            for r in range(1, 7):
                row_vals = [ws.cell(row=r, column=c).value for c in range(1, 7)]
                print(f"  Row {r}: {row_vals}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Correct department data rows with accurate counts (0.3 points)
    # Expected: 5 departments (Math, English, Biology, History, Computer Science),
    # each with 25 enrollments per course level (100, 200, 300, 400)
    # Grand total per department = 100
    try:
        expected_departments = {'math', 'english', 'biology', 'history', 'computer science'}
        expected_count_per_level = 25
        expected_row_total = 100

        # Find the header row again, then look at data rows
        # Use same logic: find row with 'department' in col 1 AND level numbers in other cols
        header_row = None
        for r in range(1, 7):
            c1_val = str(ws.cell(row=r, column=1).value or '').strip().lower()
            if 'department' not in c1_val:
                continue
            row_headers = [str(ws.cell(row=r, column=c).value or '').lower() for c in range(2, 7)]
            row_text = ' '.join(row_headers)
            if '100' in row_text or '200' in row_text or '300' in row_text or '400' in row_text:
                header_row = r
                break

        if header_row is None:
            print("FAIL: Component 3 — Could not find header row to locate data rows")
        else:
            found_departments = {}
            correct_counts = 0
            total_dept_checks = 0

            # Scan rows after header for department data
            for r in range(header_row + 1, ws.max_row + 1):
                dept_val = ws.cell(row=r, column=1).value
                if dept_val is None:
                    continue
                dept_str = str(dept_val).strip().lower()
                if dept_str in expected_departments or dept_str == 'grand total':
                    if dept_str == 'grand total':
                        continue  # Skip grand total row here
                    # Read counts for levels 100, 200, 300, 400 (columns 2-5)
                    counts = [ws.cell(row=r, column=c).value for c in range(2, 6)]
                    found_departments[dept_str] = counts

                    # Check each count
                    for cnt in counts:
                        total_dept_checks += 1
                        try:
                            if int(cnt) == expected_count_per_level:
                                correct_counts += 1
                        except (TypeError, ValueError):
                            pass

            # Check we found all 5 departments
            found_set = set(found_departments.keys())
            all_depts_found = expected_departments.issubset(found_set)

            if all_depts_found and total_dept_checks == 20 and correct_counts == 20:
                print(f"PASS: Component 3 — All 5 departments found with correct counts of 25 per level (0.3 pts)")
                total_score += 0.3
            elif all_depts_found and correct_counts >= 15:
                # Partial: departments present but some counts wrong
                print(f"FAIL: Component 3 — Departments found but counts partially incorrect. "
                      f"correct={correct_counts}/20. Found: {found_departments}")
            elif all_depts_found:
                print(f"FAIL: Component 3 — All departments found but counts are wrong. "
                      f"correct={correct_counts}/20. Expected 25 per cell. Found: {found_departments}")
            else:
                missing = expected_departments - found_set
                print(f"FAIL: Component 3 — Missing departments: {missing}. Found: {list(found_departments.keys())}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Grand total row with correct totals (0.2 points)
    # Expected: grand total row at bottom with 125 per level and 500 total
    try:
        grand_total_row = None
        for r in range(1, ws.max_row + 1):
            cell_val = ws.cell(row=r, column=1).value
            if cell_val is not None and 'grand' in str(cell_val).lower():
                grand_total_row = r
                break

        if grand_total_row is None:
            print("FAIL: Component 4 — No 'Grand Total' row found in Enrollment Summary sheet")
        else:
            # Expected: 125 for each level column (2-5) and 500 for total (col 6)
            gt_level_counts = [ws.cell(row=grand_total_row, column=c).value for c in range(2, 6)]
            gt_total = ws.cell(row=grand_total_row, column=6).value

            gt_errors = []
            for i, cnt in enumerate(gt_level_counts):
                try:
                    if int(cnt) != 125:
                        gt_errors.append(f"Level col {i+2}: expected 125, got {cnt}")
                except (TypeError, ValueError):
                    gt_errors.append(f"Level col {i+2}: non-numeric value {repr(cnt)}")

            # Check grand total of 500
            try:
                if gt_total is None or int(gt_total) != 500:
                    gt_errors.append(f"Grand total col 6: expected 500, got {repr(gt_total)}")
            except (TypeError, ValueError):
                gt_errors.append(f"Grand total col 6: non-numeric value {repr(gt_total)}")

            if not gt_errors:
                print(f"PASS: Component 4 — Grand total row correct: 125 per level, 500 total (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 4 — Grand total row errors: {gt_errors}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
