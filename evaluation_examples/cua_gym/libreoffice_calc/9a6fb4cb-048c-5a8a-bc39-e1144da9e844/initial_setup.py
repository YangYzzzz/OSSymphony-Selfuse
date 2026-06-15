"""
Initial Setup: Enrollment data for pivot table task
Task ID: calc_edu_enrollment_pivot_013
Domain: libreoffice_calc
"""

import openpyxl
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_enrollment_pivot_013'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Enrollment'

    # Headers: Student ID (A), Name (B), Department (C), Course Number (D),
    #          Course Level (E), Credits (F), Instructor (G)
    headers = ['Student ID', 'Name', 'Department', 'Course Number', 'Course Level', 'Credits', 'Instructor']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    departments = ['Math', 'English', 'Biology', 'History', 'Computer Science']

    # Course info per department per level
    course_catalog = {
        'Math': {
            100: [('MATH101', 'Calculus I'), ('MATH102', 'Pre-Calculus')],
            200: [('MATH201', 'Calculus II'), ('MATH202', 'Linear Algebra')],
            300: [('MATH301', 'Differential Equations'), ('MATH302', 'Statistics')],
            400: [('MATH401', 'Real Analysis'), ('MATH402', 'Abstract Algebra')],
        },
        'English': {
            100: [('ENGL101', 'Composition I'), ('ENGL102', 'Literature Survey')],
            200: [('ENGL201', 'Composition II'), ('ENGL202', 'World Literature')],
            300: [('ENGL301', 'Creative Writing'), ('ENGL302', 'British Literature')],
            400: [('ENGL401', 'Senior Seminar'), ('ENGL402', 'Literary Theory')],
        },
        'Biology': {
            100: [('BIOL101', 'General Biology I'), ('BIOL102', 'General Biology II')],
            200: [('BIOL201', 'Cell Biology'), ('BIOL202', 'Genetics')],
            300: [('BIOL301', 'Microbiology'), ('BIOL302', 'Ecology')],
            400: [('BIOL401', 'Molecular Biology'), ('BIOL402', 'Immunology')],
        },
        'History': {
            100: [('HIST101', 'World History I'), ('HIST102', 'World History II')],
            200: [('HIST201', 'US History I'), ('HIST202', 'US History II')],
            300: [('HIST301', 'European History'), ('HIST302', 'Asian History')],
            400: [('HIST401', 'Historical Methods'), ('HIST402', 'Senior Thesis')],
        },
        'Computer Science': {
            100: [('CS101', 'Intro to Computing'), ('CS102', 'Programming I')],
            200: [('CS201', 'Data Structures'), ('CS202', 'Algorithms')],
            300: [('CS301', 'Operating Systems'), ('CS302', 'Databases')],
            400: [('CS401', 'Machine Learning'), ('CS402', 'Software Engineering')],
        },
    }

    instructor_pool = {
        'Math': ['Dr. Patricia Walsh', 'Prof. George Navarro', 'Dr. Linda Kowalski', 'Prof. Samuel Okafor'],
        'English': ['Prof. Diana Fernandez', 'Dr. Robert Ashford', 'Prof. Mei-Ling Chung', 'Dr. Anthony Rivers'],
        'Biology': ['Dr. Karen Summers', 'Prof. James Obasi', 'Dr. Nadia Petrov', 'Prof. Carlos Mendez'],
        'History': ['Dr. Eleanor Burke', 'Prof. Winston Adeyemi', 'Dr. Sophia Larsson', 'Prof. Derek Yamamoto'],
        'Computer Science': ['Dr. Rachel Kim', 'Prof. Marcus Thompson', 'Dr. Aisha Patel', 'Prof. Ben Nakamura'],
    }

    first_names = [
        'Emma', 'Liam', 'Olivia', 'Noah', 'Ava', 'William', 'Sophia', 'James',
        'Isabella', 'Oliver', 'Mia', 'Benjamin', 'Charlotte', 'Elijah', 'Amelia',
        'Lucas', 'Harper', 'Mason', 'Evelyn', 'Logan', 'Abigail', 'Ethan', 'Emily',
        'Aiden', 'Elizabeth', 'Jackson', 'Mila', 'Sebastian', 'Ella', 'Mateo',
        'Avery', 'Jack', 'Sofia', 'Owen', 'Camila', 'Wyatt', 'Aria', 'Charlie',
        'Scarlett', 'Chloe', 'Joseph', 'Victoria', 'Henry', 'Madison', 'Samuel',
        'Luna', 'David', 'Grace', 'Wyatt', 'Naomi'
    ]
    last_names = [
        'Smith', 'Johnson', 'Williams', 'Brown', 'Jones', 'Garcia', 'Miller',
        'Davis', 'Rodriguez', 'Martinez', 'Hernandez', 'Lopez', 'Gonzalez',
        'Wilson', 'Anderson', 'Thomas', 'Taylor', 'Moore', 'Jackson', 'Martin',
        'Lee', 'Perez', 'Thompson', 'White', 'Harris', 'Sanchez', 'Clark',
        'Ramirez', 'Lewis', 'Robinson', 'Walker', 'Young', 'Allen', 'King',
        'Wright', 'Scott', 'Torres', 'Nguyen', 'Hill', 'Flores', 'Green',
        'Adams', 'Nelson', 'Baker', 'Hall', 'Rivera', 'Campbell', 'Mitchell',
        'Carter', 'Roberts'
    ]

    credits_map = {100: 3, 200: 3, 300: 3, 400: 4}

    # Build 500 enrollment records
    # Distribute evenly: 500 records / 5 departments = 100 per dept
    # Per dept, 4 levels * 25 = 100 per dept
    records = []
    student_id = 1001
    for dept in departments:
        for level in [100, 200, 300, 400]:
            count = 25  # 25 students per dept per level = 500 total
            courses = course_catalog[dept][level]
            instructors = instructor_pool[dept]
            for i in range(count):
                fname = random.choice(first_names)
                lname = random.choice(last_names)
                name = f'{fname} {lname}'
                course_code, _ = random.choice(courses)
                instructor = random.choice(instructors)
                credits = credits_map[level]
                records.append([
                    f'STU{student_id:04d}',
                    name,
                    dept,
                    course_code,
                    level,
                    credits,
                    instructor
                ])
                student_id += 1

    # Shuffle to mix up departments/levels
    random.shuffle(records)

    for r, row_data in enumerate(records, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 24

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Rows: {ws.max_row - 1} data rows (plus header)')
    print(f'Columns: {headers}')

create_initial()
