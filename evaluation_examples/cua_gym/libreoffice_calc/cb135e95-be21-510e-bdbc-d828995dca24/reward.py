"""
Reward Script: Export workbook to PDF with all sheets included
Task ID: calc_gsi_030
Domain: libreoffice_calc
Scoring:
  Component 1: PDF file exists and is non-trivial (0.2 points)
  Component 2: PDF has enough pages to contain all 6 sheets (0.3 points)
  Component 3: PDF contains data from all 6 monthly sheets (0.5 points)
"""

import os

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_030'


def persist_app_state(domain):
    """Try to save any unsaved state in LibreOffice."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task():
    """
    Verify that the workbook was exported to PDF with all 6 sheets included.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    pdf_path = f'{WORKDIR}/{TASK_ID}.pdf'
    xlsx_path = f'{WORKDIR}/{TASK_ID}.xlsx'

    # Component 1: PDF file exists and is non-trivial (0.2 points)
    try:
        if os.path.exists(pdf_path):
            size = os.path.getsize(pdf_path)
            if size > 5000:
                print(f"PASS: Component 1 — PDF exists at {pdf_path}, size={size} bytes (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 1 — PDF exists but too small ({size} bytes), likely empty/corrupt")
        else:
            print(f"FAIL: Component 1 — PDF not found at {pdf_path}")
            # No PDF means no further checks possible
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: PDF has enough pages to represent all 6 sheets (0.3 points)
    # The workbook has 6 sheets with ~12 rows and 10 columns each.
    # All 6 sheets should produce >= 6 pages total.
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(pdf_path)
        page_count = len(reader.pages)
        if page_count >= 6:
            print(f"PASS: Component 2 — PDF has {page_count} pages (>= 6, consistent with all sheets) (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — PDF has only {page_count} pages (expected >= 6 for 6 sheets)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: PDF contains data from all 6 monthly sheets (0.5 points)
    # Verify by extracting unique revenue values from each sheet in the xlsx
    # and checking they appear in the PDF text.
    try:
        import openpyxl
        from PyPDF2 import PdfReader

        # Extract all text from PDF
        reader = PdfReader(pdf_path)
        pdf_text = ""
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                pdf_text += page_text + "\n"

        # Load xlsx to get unique identifying values from each sheet
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        expected_sheets = wb.sheetnames
        sheets_found = 0

        for sheet_name in expected_sheets:
            ws = wb[sheet_name]
            # Get a distinctive value from this sheet (Engineering Revenue - B2)
            val = ws.cell(row=2, column=2).value
            if val is not None:
                # Try multiple format representations
                formatted_comma = f"{val:,.2f}"
                formatted_plain = f"{val:.2f}"
                formatted_int = str(int(val)) if val == int(val) else None

                if formatted_comma in pdf_text or formatted_plain in pdf_text:
                    sheets_found += 1
                    print(f"  Sheet '{sheet_name}': found distinctive value ${formatted_comma} in PDF")
                elif formatted_int and formatted_int in pdf_text:
                    sheets_found += 1
                    print(f"  Sheet '{sheet_name}': found value {formatted_int} in PDF")
                else:
                    # Try removing commas from PDF text and matching
                    pdf_no_comma = pdf_text.replace(",", "")
                    if formatted_plain in pdf_no_comma:
                        sheets_found += 1
                        print(f"  Sheet '{sheet_name}': found value {formatted_plain} in PDF (no-comma match)")
                    else:
                        print(f"  Sheet '{sheet_name}': value ${formatted_comma} NOT found in PDF")
            else:
                print(f"  Sheet '{sheet_name}': no cached value in B2, skipping")

        if sheets_found == len(expected_sheets):
            print(f"PASS: Component 3 — All {sheets_found}/{len(expected_sheets)} sheets found in PDF (0.5 pts)")
            total_score += 0.5
        elif sheets_found > 1:
            partial = round(0.5 * (sheets_found / len(expected_sheets)), 2)
            if partial > 0:
                print(f"PARTIAL: Component 3 — {sheets_found}/{len(expected_sheets)} sheets found in PDF ({partial} pts)")
                total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {sheets_found}/{len(expected_sheets)} sheets found in PDF")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")
verify_task()
