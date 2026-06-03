"""
Initial Setup: Export workbook to PDF with all sheets included
Task ID: calc_gsi_030
Domain: libreoffice_calc

Creates a 6-sheet monthly report workbook and opens it in LibreOffice Calc.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_030'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # Common styles
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    currency_fmt = '$#,##0.00'
    pct_fmt = '0.0%'
    date_fmt = 'yyyy-mm-dd'
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    months = ["January", "February", "March", "April", "May", "June"]
    departments = [
        "Engineering", "Marketing", "Sales", "Human Resources",
        "Finance", "Operations", "Customer Support", "Product",
        "Legal", "Research & Development"
    ]

    # Revenue base values per department (vary per month)
    base_revenues = [185000, 92000, 245000, 68000, 55000, 132000, 78000, 115000, 42000, 165000]
    base_expenses = [152000, 78000, 118000, 62000, 48000, 105000, 65000, 98000, 38000, 140000]
    headcounts = [45, 18, 32, 12, 10, 28, 22, 15, 8, 35]

    for idx, month in enumerate(months):
        if idx == 0:
            ws = wb.active
            ws.title = month
        else:
            ws = wb.create_sheet(month)

        # Headers
        headers = ["Department", "Revenue", "Expenses", "Net Profit",
                    "Headcount", "Revenue per Head", "Profit Margin",
                    "YoY Growth", "Budget Variance", "Status"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Column widths
        col_widths = {"A": 22, "B": 15, "C": 15, "D": 15, "E": 12,
                      "F": 18, "G": 14, "H": 12, "I": 16, "J": 14}
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Data rows
        import random
        random.seed(42 + idx)  # Deterministic but different per month

        statuses = ["On Track", "On Track", "At Risk", "On Track", "Ahead",
                     "On Track", "On Track", "Ahead", "On Track", "On Track"]

        for r, dept in enumerate(departments, 2):
            dept_idx = r - 2
            # Vary revenues slightly per month
            month_factor = 1.0 + (idx * 0.02) + random.uniform(-0.05, 0.05)
            revenue = round(base_revenues[dept_idx] * month_factor, 2)
            expenses = round(base_expenses[dept_idx] * (1.0 + random.uniform(-0.03, 0.03)), 2)
            net_profit = round(revenue - expenses, 2)
            hc = headcounts[dept_idx] + random.randint(-1, 2)
            rev_per_head = round(revenue / hc, 2) if hc > 0 else 0
            profit_margin = round(net_profit / revenue, 4) if revenue > 0 else 0
            yoy_growth = round(random.uniform(0.02, 0.18), 4)
            budget_var = round(random.uniform(-0.08, 0.12), 4)

            ws.cell(row=r, column=1, value=dept).border = thin_border
            ws.cell(row=r, column=2, value=revenue).border = thin_border
            ws.cell(row=r, column=2).number_format = currency_fmt
            ws.cell(row=r, column=3, value=expenses).border = thin_border
            ws.cell(row=r, column=3).number_format = currency_fmt
            ws.cell(row=r, column=4, value=net_profit).border = thin_border
            ws.cell(row=r, column=4).number_format = currency_fmt
            ws.cell(row=r, column=5, value=hc).border = thin_border
            ws.cell(row=r, column=6, value=rev_per_head).border = thin_border
            ws.cell(row=r, column=6).number_format = currency_fmt
            ws.cell(row=r, column=7, value=profit_margin).border = thin_border
            ws.cell(row=r, column=7).number_format = pct_fmt
            ws.cell(row=r, column=8, value=yoy_growth).border = thin_border
            ws.cell(row=r, column=8).number_format = pct_fmt
            ws.cell(row=r, column=9, value=budget_var).border = thin_border
            ws.cell(row=r, column=9).number_format = pct_fmt
            ws.cell(row=r, column=10, value=statuses[dept_idx]).border = thin_border

        # Totals row
        total_row = len(departments) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=1).border = thin_border
        for col in range(2, 10):
            cell = ws.cell(row=total_row, column=col)
            cell.border = thin_border
            if col in (2, 3, 4, 6):
                cell.number_format = currency_fmt
            elif col in (7, 8, 9):
                cell.number_format = pct_fmt

        # Calculate totals for numeric columns
        total_revenue = sum(ws.cell(row=r, column=2).value for r in range(2, total_row))
        total_expenses = sum(ws.cell(row=r, column=3).value for r in range(2, total_row))
        total_profit = sum(ws.cell(row=r, column=4).value for r in range(2, total_row))
        total_hc = sum(ws.cell(row=r, column=5).value for r in range(2, total_row))

        ws.cell(row=total_row, column=2, value=round(total_revenue, 2))
        ws.cell(row=total_row, column=3, value=round(total_expenses, 2))
        ws.cell(row=total_row, column=4, value=round(total_profit, 2))
        ws.cell(row=total_row, column=5, value=total_hc)
        if total_hc > 0:
            ws.cell(row=total_row, column=6, value=round(total_revenue / total_hc, 2))
        if total_revenue > 0:
            ws.cell(row=total_row, column=7, value=round(total_profit / total_revenue, 4))

        # Freeze header row
        ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
