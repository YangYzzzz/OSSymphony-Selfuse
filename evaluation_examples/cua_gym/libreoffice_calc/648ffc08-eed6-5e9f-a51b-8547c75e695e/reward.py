"""
Reward Script: Read audit_checklist.docx and implement data validation, cross-sheet
formulas, error-check sheet, and conditional formatting in financial_model.xlsx.
Task ID: osworld_multi_apps_docx_to_calc_011
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Inputs sheet has scenario dropdown data validation on B3
  Component 2 (0.25): EBITDA sheet has cross-sheet formula references in Total Revenue/Costs rows
  Component 3 (0.25): EBITDA sheet has EBITDA calculation formulas (Revenue - Costs) in row 6
  Component 4 (0.25): Error_Check sheet exists with IF deviation formulas and conditional formatting
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_docx_to_calc_011'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # -------------------------------------------------------------------------
    # Component 1: Inputs sheet has dropdown data validation for scenario
    #   selection (0.25 points)
    # The task requires "input validation (dropdowns for scenario selection)".
    # In the golden file, cell B3 in Inputs has a list validation:
    #   "Base Case,Upside,Downside"
    # -------------------------------------------------------------------------
    try:
        if 'Inputs' not in wb.sheetnames:
            print("FAIL: Component 1 — 'Inputs' sheet not found")
        else:
            ws_inputs = wb['Inputs']
            dvs = ws_inputs.data_validations.dataValidation
            # Look for a list-type data validation covering B3 or nearby scenario cell
            found_scenario_dv = False
            for dv in dvs:
                if dv.type == 'list':
                    sqref_str = str(dv.sqref)
                    formula_str = str(dv.formula1) if dv.formula1 else ''
                    # Check if it applies to B3 (scenario cell) or contains scenario options
                    if 'B3' in sqref_str or any(
                        kw in formula_str for kw in ['Base Case', 'Upside', 'Downside']
                    ):
                        found_scenario_dv = True
                        print(
                            f"PASS: Component 1 — Scenario dropdown found: "
                            f"type={dv.type}, sqref={sqref_str}, formula1={formula_str} (0.25 pts)"
                        )
                        break
            if not found_scenario_dv:
                print(
                    f"FAIL: Component 1 — No scenario dropdown data validation found on Inputs sheet "
                    f"(found {len(dvs)} validations total)"
                )
                for dv in dvs:
                    print(f"  DV: type={dv.type}, sqref={dv.sqref}, formula1={dv.formula1}")
            else:
                total_score += 0.25
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: EBITDA sheet has cross-sheet formula references for
    #   Total Revenue and Total Costs rows (0.25 points)
    # The task requires "cross-sheet formula checks (EBITDA = Revenue - Costs)".
    # In the golden file, EBITDA rows 4 and 5 use =Revenue!... and =Costs!...
    # references.
    # -------------------------------------------------------------------------
    try:
        if 'EBITDA' not in wb.sheetnames:
            print("FAIL: Component 2 — 'EBITDA' sheet not found")
        else:
            ws_ebitda = wb['EBITDA']
            cross_sheet_refs_found = 0
            # Scan EBITDA sheet for cross-sheet formula references
            for row in ws_ebitda.iter_rows(values_only=False):
                for cell in row:
                    val = cell.value
                    if isinstance(val, str) and val.startswith('='):
                        if 'Revenue!' in val or 'Costs!' in val:
                            cross_sheet_refs_found += 1

            if cross_sheet_refs_found >= 2:
                print(
                    f"PASS: Component 2 — Cross-sheet formula references found in EBITDA sheet "
                    f"({cross_sheet_refs_found} references) (0.25 pts)"
                )
                total_score += 0.25
            else:
                print(
                    f"FAIL: Component 2 — Expected >=2 cross-sheet references (Revenue!/Costs!) "
                    f"in EBITDA sheet, found {cross_sheet_refs_found}"
                )
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: EBITDA sheet has EBITDA calculation formulas in the EBITDA
    #   row (Revenue - Costs pattern) (0.25 points)
    # The task requires cross-sheet formula checks (EBITDA = Revenue - Costs).
    # In the golden file, row 6 (EBITDA row) uses =B4-B5 style subtraction formulas.
    # -------------------------------------------------------------------------
    try:
        if 'EBITDA' not in wb.sheetnames:
            print("FAIL: Component 3 — 'EBITDA' sheet not found")
        else:
            ws_ebitda = wb['EBITDA']
            ebitda_formula_found = False
            # Look for a row labelled EBITDA with subtraction formulas
            for row in ws_ebitda.iter_rows(values_only=False):
                # Check if this row has EBITDA label in column A
                label_cell = row[0]
                if label_cell.value and 'EBITDA' in str(label_cell.value).upper():
                    # Check if any numeric column in this row has a subtraction formula
                    for cell in row[1:]:
                        val = cell.value
                        if isinstance(val, str) and val.startswith('=') and '-' in val:
                            ebitda_formula_found = True
                            print(
                                f"PASS: Component 3 — EBITDA formula found at {cell.coordinate}: "
                                f"{val} (0.25 pts)"
                            )
                            break
                    if ebitda_formula_found:
                        break

            if not ebitda_formula_found:
                print(
                    "FAIL: Component 3 — No EBITDA subtraction formula (Revenue - Costs) found "
                    "in EBITDA sheet row labelled 'EBITDA'"
                )
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")
    else:
        if ebitda_formula_found:
            total_score += 0.25

    # -------------------------------------------------------------------------
    # Component 4: Error_Check sheet exists with IF formulas flagging deviations
    #   >5% and conditional formatting highlighting discrepancies (0.25 points)
    # The task requires "an error-check sheet with IF formulas flagging any cell
    # where computed value deviates more than 5% from manual entry, and conditional
    # formatting highlighting discrepancies in red."
    # -------------------------------------------------------------------------
    try:
        if 'Error_Check' not in wb.sheetnames:
            print("FAIL: Component 4 — 'Error_Check' sheet not found")
        else:
            ws_ec = wb['Error_Check']
            # Check for IF formulas in the Status column
            if_formula_count = 0
            for row in ws_ec.iter_rows(values_only=False):
                for cell in row:
                    val = cell.value
                    if isinstance(val, str) and val.upper().startswith('=IF('):
                        if_formula_count += 1

            # Check for conditional formatting on Error_Check or EBITDA sheet
            ec_has_cf = len(list(ws_ec.conditional_formatting._cf_rules.keys())) > 0

            # Also accept CF on EBITDA sheet as "discrepancy highlighting"
            ebitda_has_cf = False
            if 'EBITDA' in wb.sheetnames:
                ws_e = wb['EBITDA']
                ebitda_has_cf = len(list(ws_e.conditional_formatting._cf_rules.keys())) > 0

            has_cf = ec_has_cf or ebitda_has_cf

            if if_formula_count >= 1 and has_cf:
                print(
                    f"PASS: Component 4 — Error_Check sheet found with {if_formula_count} IF "
                    f"formula(s) and conditional formatting (EC_CF={ec_has_cf}, EBITDA_CF={ebitda_has_cf}) "
                    f"(0.25 pts)"
                )
                total_score += 0.25
            elif if_formula_count >= 1 and not has_cf:
                print(
                    f"PARTIAL: Component 4 — Error_Check sheet has IF formulas "
                    f"({if_formula_count}) but no conditional formatting found"
                )
            elif if_formula_count == 0 and has_cf:
                print(
                    "PARTIAL: Component 4 — Conditional formatting found but no IF formulas "
                    "in Error_Check sheet"
                )
            else:
                print(
                    "FAIL: Component 4 — Error_Check sheet exists but missing IF formulas "
                    f"({if_formula_count}) and/or conditional formatting (CF={has_cf})"
                )
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path on the VM
file_path = f'{WORKDIR}/financial_model.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
