"""
Initial Setup: Monthly Utility Bill Tracker
Task ID: calc_grs_030
Domain: libreoffice_calc

Creates a spreadsheet with 6 utility types and 12 months of realistic cost data.
No formulas, no conditional formatting, no charts - those are the task for the agent.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_030'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Utility Bills"

    # --- Row 1: Title ---
    ws.merge_cells("A1:P1")
    ws["A1"] = "Monthly Utility Bill Tracker 2025"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # --- Row 2: Headers ---
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    headers = ["Utility Type"] + months + ["Annual Total", "12-Month Average", "YoY Change %"]

    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2E75B6", end_color="FF2E75B6", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Rows 3-8: Utility data (6 types) ---
    # Realistic monthly utility costs with summer electricity spike
    utility_data = {
        "Electricity": [128.45, 115.20, 109.75, 132.60, 168.90, 215.40,
                        258.30, 267.15, 198.50, 145.80, 118.60, 135.25],
        "Gas": [142.30, 138.50, 112.80, 78.60, 45.20, 32.10,
                28.50, 29.80, 38.60, 72.40, 118.90, 148.50],
        "Water/Sewer": [65.80, 62.40, 68.90, 74.20, 82.50, 95.30,
                        108.40, 112.60, 98.70, 78.40, 66.20, 63.80],
        "Internet": [79.99, 79.99, 79.99, 79.99, 79.99, 79.99,
                     79.99, 79.99, 79.99, 79.99, 79.99, 79.99],
        "Phone": [85.00, 85.00, 87.50, 85.00, 85.00, 85.00,
                  92.30, 85.00, 85.00, 85.00, 85.00, 85.00],
        "Trash": [35.00, 35.00, 35.00, 35.00, 35.00, 35.00,
                  35.00, 35.00, 35.00, 35.00, 35.00, 35.00],
    }

    data_font = Font(name="Arial", size=10)
    currency_fmt = '$#,##0.00'
    label_font = Font(name="Arial", size=10, bold=True)

    thin_side = Side(style="thin", color="D9D9D9")
    data_border = Border(left=thin_side, right=thin_side,
                         top=thin_side, bottom=thin_side)

    row_idx = 3
    for utility_name, monthly_values in utility_data.items():
        # Utility name in column A
        cell_a = ws.cell(row=row_idx, column=1, value=utility_name)
        cell_a.font = label_font
        cell_a.border = data_border

        # Monthly values in columns B-M (2-13)
        for col_offset, val in enumerate(monthly_values):
            cell = ws.cell(row=row_idx, column=col_offset + 2, value=val)
            cell.number_format = currency_fmt
            cell.font = data_font
            cell.alignment = Alignment(horizontal="right")
            cell.border = data_border

        # Leave columns N, O, P (Annual Total, Average, YoY Change) EMPTY
        # These are what the agent needs to fill with formulas
        for empty_col in [14, 15, 16]:
            cell = ws.cell(row=row_idx, column=empty_col)
            cell.border = data_border

        row_idx += 1

    # --- Row 9: Monthly Total row ---
    # Label only - no formulas (agent must add them)
    ws.cell(row=9, column=1, value="Monthly Total").font = Font(
        name="Arial", size=10, bold=True, color="1F4E79"
    )
    total_fill = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")
    for col in range(1, 17):
        cell = ws.cell(row=9, column=col)
        cell.fill = total_fill
        cell.border = data_border
        if col >= 2:
            cell.number_format = currency_fmt

    # --- Column widths ---
    ws.column_dimensions["A"].width = 18
    for col_letter in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]:
        ws.column_dimensions[col_letter].width = 12
    ws.column_dimensions["N"].width = 14
    ws.column_dimensions["O"].width = 16
    ws.column_dimensions["P"].width = 16

    # --- Row heights ---
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 25

    # Freeze header rows
    ws.freeze_panes = "B3"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
