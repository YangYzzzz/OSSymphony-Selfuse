"""
Reward Script: Monthly Utility Bill Tracker
Task ID: calc_grs_030
Domain: libreoffice_calc
Scoring:
  Component 1: Annual Total SUM formulas in N3:N8 (0.20 pts)
  Component 2: 12-Month Average formulas in O3:O8 (0.15 pts)
  Component 3: Monthly Total SUM formulas in B9:M9 (0.20 pts)
  Component 4: Grand total / overall average in N9, O9 (0.10 pts)
  Component 5: Conditional formatting rules for spike detection (0.15 pts)
  Component 6: Stacked area chart present (0.10 pts)
  Component 7: Horizontal bar chart present (0.10 pts)
"""

import os
import openpyxl
from openpyxl.chart import AreaChart, BarChart

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_030'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Find the sheet (should be 'Utility Bills' or first sheet)
    ws = None
    for name in wb.sheetnames:
        if 'util' in name.lower() or 'bill' in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb.worksheets[0]

    # ---------------------------------------------------------------
    # Component 1: Annual Total SUM formulas in N3:N8 (0.20 points)
    # Each utility row should have =SUM(B_:M_) in column N
    # This is EMPTY in initial_env, so only scores task-introduced change.
    # ---------------------------------------------------------------
    try:
        sum_formula_count = 0
        for row in range(3, 9):  # rows 3-8
            val = ws.cell(row=row, column=14).value  # column N
            if val is not None and isinstance(val, str) and 'SUM' in val.upper():
                # Check it references the correct row range (B to M)
                expected_part = f'B{row}'
                if expected_part.upper() in val.upper().replace(' ', ''):
                    sum_formula_count += 1
        if sum_formula_count == 6:
            print(f"PASS: Component 1 — All 6 Annual Total SUM formulas present in N3:N8 (0.20 pts)")
            total_score += 0.20
        elif sum_formula_count >= 3:
            partial = round(0.20 * sum_formula_count / 6, 2)
            print(f"PARTIAL: Component 1 — {sum_formula_count}/6 Annual Total SUM formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {sum_formula_count}/6 Annual Total SUM formulas found in N3:N8")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ---------------------------------------------------------------
    # Component 2: 12-Month Average formulas in O3:O8 (0.15 points)
    # Each utility row should have =AVERAGE(B_:M_) in column O
    # This is EMPTY in initial_env.
    # ---------------------------------------------------------------
    try:
        avg_formula_count = 0
        for row in range(3, 9):
            val = ws.cell(row=row, column=15).value  # column O
            if val is not None and isinstance(val, str) and 'AVERAGE' in val.upper():
                expected_part = f'B{row}'
                if expected_part.upper() in val.upper().replace(' ', ''):
                    avg_formula_count += 1
        if avg_formula_count == 6:
            print(f"PASS: Component 2 — All 6 Average formulas present in O3:O8 (0.15 pts)")
            total_score += 0.15
        elif avg_formula_count >= 3:
            partial = round(0.15 * avg_formula_count / 6, 2)
            print(f"PARTIAL: Component 2 — {avg_formula_count}/6 Average formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {avg_formula_count}/6 Average formulas found in O3:O8")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---------------------------------------------------------------
    # Component 3: Monthly Total SUM formulas in row 9, B9:M9 (0.20 points)
    # Each month column should have =SUM(B3:B8) style formula
    # Row 9 is EMPTY in initial_env.
    # ---------------------------------------------------------------
    try:
        monthly_total_count = 0
        for col in range(2, 14):  # columns B(2) through M(13)
            val = ws.cell(row=9, column=col).value
            if val is not None and isinstance(val, str) and 'SUM' in val.upper():
                # Should reference rows 3 through 8 in the same column
                col_letter = openpyxl.utils.get_column_letter(col)
                if f'{col_letter}3' in val.upper().replace(' ', '') or f'{col_letter.lower()}3' in val.lower().replace(' ', ''):
                    monthly_total_count += 1
        if monthly_total_count == 12:
            print(f"PASS: Component 3 — All 12 Monthly Total SUM formulas present in B9:M9 (0.20 pts)")
            total_score += 0.20
        elif monthly_total_count >= 6:
            partial = round(0.20 * monthly_total_count / 12, 2)
            print(f"PARTIAL: Component 3 — {monthly_total_count}/12 Monthly Total SUM formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {monthly_total_count}/12 Monthly Total SUM formulas found in B9:M9")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ---------------------------------------------------------------
    # Component 4: Grand total (N9) and overall average (O9) (0.10 points)
    # N9 should have a SUM formula, O9 should have an AVERAGE formula.
    # Both are EMPTY in initial_env.
    # ---------------------------------------------------------------
    try:
        n9_val = ws.cell(row=9, column=14).value  # N9
        o9_val = ws.cell(row=9, column=15).value  # O9
        sub_score = 0.0
        n9_ok = (n9_val is not None and isinstance(n9_val, str) and 'SUM' in n9_val.upper())
        o9_ok = (o9_val is not None and isinstance(o9_val, str) and 'AVERAGE' in o9_val.upper())
        if n9_ok:
            sub_score += 0.05
        if o9_ok:
            sub_score += 0.05
        if sub_score == 0.10:
            print(f"PASS: Component 4 — Grand total (N9={n9_val}) and overall average (O9={o9_val}) present (0.10 pts)")
            total_score += 0.10
        elif sub_score > 0:
            print(f"PARTIAL: Component 4 — N9={'OK' if n9_ok else 'MISSING'}, O9={'OK' if o9_ok else 'MISSING'} ({sub_score} pts)")
            total_score += sub_score
        else:
            print(f"FAIL: Component 4 — N9={n9_val}, O9={o9_val} — expected SUM and AVERAGE formulas")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ---------------------------------------------------------------
    # Component 5: Conditional formatting rules for spike detection (0.15 points)
    # Should have conditional formatting on utility data rows (rows 3-8)
    # that highlights when a month is >20% above that utility's average.
    # There are 0 conditional formatting rules in initial_env.
    # ---------------------------------------------------------------
    try:
        cf_rules = list(ws.conditional_formatting)
        # Count how many utility rows (3-8) have a conditional formatting rule
        rows_with_cf = set()
        avg_formula_count = 0
        for cf in cf_rules:
            for rule in cf.rules:
                # Check if rule formula references AVERAGE (spike detection)
                if rule.formula:
                    for f in rule.formula:
                        if 'AVERAGE' in str(f).upper() and '1.2' in str(f):
                            avg_formula_count += 1
                # Track which rows are covered
                range_str = str(cf)
                # Extract row numbers from the range string
                for row in range(3, 9):
                    if str(row) in range_str:
                        rows_with_cf.add(row)

        if len(rows_with_cf) >= 6 and avg_formula_count > 0:
            print(f"PASS: Component 5 — Conditional formatting with AVERAGE-based spike detection on {len(rows_with_cf)} utility rows (0.15 pts)")
            total_score += 0.15
        elif len(rows_with_cf) >= 3 and avg_formula_count > 0:
            partial = round(0.15 * len(rows_with_cf) / 6, 2)
            print(f"PARTIAL: Component 5 — Conditional formatting on {len(rows_with_cf)}/6 utility rows ({partial} pts)")
            total_score += partial
        elif len(cf_rules) > 0:
            # Some conditional formatting exists but might not match exact pattern
            print(f"PARTIAL: Component 5 — {len(cf_rules)} conditional formatting rules found but pattern incomplete (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — No conditional formatting rules found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # ---------------------------------------------------------------
    # Component 6: Stacked area chart for monthly cost trend (0.10 points)
    # No charts exist in initial_env.
    # ---------------------------------------------------------------
    try:
        charts = ws._charts
        area_charts = [c for c in charts if isinstance(c, AreaChart)]
        stacked_area = [c for c in area_charts if getattr(c, 'grouping', None) == 'stacked']
        if len(stacked_area) >= 1:
            print(f"PASS: Component 6 — Stacked area chart found (0.10 pts)")
            total_score += 0.10
        elif len(area_charts) >= 1:
            print(f"PARTIAL: Component 6 — Area chart found but grouping={getattr(area_charts[0], 'grouping', 'unknown')}, expected 'stacked' (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 — No area chart found (found {len(charts)} charts total)")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # ---------------------------------------------------------------
    # Component 7: Horizontal bar chart for annual totals (0.10 points)
    # No charts exist in initial_env.
    # ---------------------------------------------------------------
    try:
        charts = ws._charts
        bar_charts = [c for c in charts if isinstance(c, BarChart)]
        horiz_bars = [c for c in bar_charts if getattr(c, 'type', None) == 'bar']
        if len(horiz_bars) >= 1:
            print(f"PASS: Component 7 — Horizontal bar chart found (0.10 pts)")
            total_score += 0.10
        elif len(bar_charts) >= 1:
            print(f"PARTIAL: Component 7 — Bar chart found but type={getattr(bar_charts[0], 'type', 'unknown')}, expected 'bar' for horizontal (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 7 — No bar chart found (found {len(charts)} charts total)")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
