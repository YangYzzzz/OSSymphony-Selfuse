"""
Reward Script: Dynamic Report Card System
Task ID: calc_wf_022
Domain: libreoffice_calc
Scoring:
  Component 1: Subject percentage formulas in Report C4:H23 (0.25)
  Component 2: Overall percentage + GPA + Grade + Pass/Fail formulas in I-L (0.25)
  Component 3: RANK formulas in M4:M23 (0.10)
  Component 4: Conditional formatting on percentage cols <40 in red (0.15)
  Component 5: Data validation on Marks sheet (0.15)
  Component 6: Print layout (landscape orientation) (0.10)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_022'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: required sheets exist
    required_sheets = ['Subjects', 'Marks', 'Report']
    for s in required_sheets:
        if s not in wb.sheetnames:
            print(f"CRITICAL: Required sheet '{s}' missing")
            print("REWARD: 0.0")
            return 0.0

    ws_report = wb['Report']
    ws_marks = wb['Marks']

    # Component 1: Subject percentage formulas in Report C4:H23 (0.25 points)
    # In initial_env these cells are None; in golden_env they have formulas like
    # =Marks!C2/Subjects!$D$2*100
    try:
        formula_count = 0
        total_cells = 0
        for row in ws_report.iter_rows(min_row=4, max_row=23, min_col=3, max_col=8):
            for cell in row:
                total_cells += 1
                val = cell.value
                if val is not None and isinstance(val, str) and val.startswith('='):
                    # Check it references Marks sheet and computes a percentage
                    val_upper = val.upper()
                    if 'MARKS!' in val_upper and ('/' in val or '*' in val):
                        formula_count += 1
        ratio = formula_count / total_cells if total_cells > 0 else 0
        if ratio >= 0.9:
            print(f"PASS: Component 1 — {formula_count}/{total_cells} subject percentage formulas found (0.25 pts)")
            total_score += 0.25
        elif ratio >= 0.5:
            partial = 0.25 * ratio
            print(f"PARTIAL: Component 1 — {formula_count}/{total_cells} formulas ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {formula_count}/{total_cells} subject percentage formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Overall %, GPA, Grade, Pass/Fail formulas in cols I-L (0.25 points)
    # I: overall percentage, J: GPA (nested IF), K: Grade (nested IF), L: Pass/Fail (AND condition)
    try:
        col_checks = {
            'I': {'name': 'Overall %', 'keywords': ['MARKS!', '/'], 'count': 0},
            'J': {'name': 'GPA', 'keywords': ['IF('], 'count': 0},
            'K': {'name': 'Grade', 'keywords': ['IF('], 'count': 0},
            'L': {'name': 'Pass/Fail', 'keywords': ['AND('], 'count': 0},
        }
        col_map = {'I': 9, 'J': 10, 'K': 11, 'L': 12}
        rows_expected = 20  # rows 4-23

        for col_letter, info in col_checks.items():
            col_idx = col_map[col_letter]
            for row_num in range(4, 24):
                val = ws_report.cell(row=row_num, column=col_idx).value
                if val is not None and isinstance(val, str) and val.startswith('='):
                    val_upper = val.upper()
                    if all(kw.upper() in val_upper for kw in info['keywords']):
                        info['count'] += 1

        passed_cols = sum(1 for info in col_checks.values() if info['count'] >= 18)
        comp2_score = 0.0
        if passed_cols == 4:
            comp2_score = 0.25
            print(f"PASS: Component 2 — All 4 formula columns (I-L) populated correctly (0.25 pts)")
        elif passed_cols > 0:
            comp2_score = 0.25 * (passed_cols / 4)
            details = "; ".join(f"{k}:{v['count']}/20" for k, v in col_checks.items())
            print(f"PARTIAL: Component 2 — {passed_cols}/4 columns correct ({comp2_score:.2f} pts) [{details}]")
        else:
            details = "; ".join(f"{k}:{v['count']}/20" for k, v in col_checks.items())
            print(f"FAIL: Component 2 — No formula columns correctly populated [{details}]")
        if comp2_score > 0:
            total_score += comp2_score
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: RANK formulas in M4:M23 (0.10 points)
    try:
        rank_count = 0
        for row_num in range(4, 24):
            val = ws_report.cell(row=row_num, column=13).value  # col M
            if val is not None and isinstance(val, str) and val.startswith('='):
                if 'RANK(' in val.upper() or 'RANK.EQ(' in val.upper() or 'RANK.AVG(' in val.upper():
                    rank_count += 1
        if rank_count >= 18:
            print(f"PASS: Component 3 — {rank_count}/20 RANK formulas found in col M (0.10 pts)")
            total_score += 0.10
        elif rank_count > 0:
            partial = 0.10 * (rank_count / 20)
            print(f"PARTIAL: Component 3 — {rank_count}/20 RANK formulas ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No RANK formulas found in col M")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Conditional formatting on percentage columns <40 in red (0.15 points)
    # Golden has CF rules on C4:C23 through H4:H23, type=cellIs, operator=lessThan, formula=['40']
    try:
        cf_rules = ws_report.conditional_formatting
        cf_cols_covered = set()
        for cf in cf_rules:
            for rule in cf.rules:
                rule_type = getattr(rule, 'type', '')
                rule_operator = getattr(rule, 'operator', '')
                rule_formula = getattr(rule, 'formula', [])
                # Check for cellIs lessThan 40
                is_less_than_40 = (
                    rule_type == 'cellIs' and
                    rule_operator == 'lessThan' and
                    rule_formula and '40' in str(rule_formula[0])
                )
                if is_less_than_40:
                    # Check for red fill
                    has_red_fill = False
                    if hasattr(rule, 'dxf') and rule.dxf:
                        if rule.dxf.fill and hasattr(rule.dxf.fill, 'fgColor'):
                            rgb = getattr(rule.dxf.fill.fgColor, 'rgb', '')
                            has_red_fill = (rgb is not None and 'FF0000' in str(rgb).upper())
                        if rule.dxf.font and hasattr(rule.dxf.font, 'color'):
                            # Also accept red font color as "highlighting in red"
                            pass
                    if has_red_fill:
                        # Parse the range to find which columns are covered
                        range_str = str(cf)
                        # Extract column letters from range like "C4:C23"
                        import re
                        matches = re.findall(r'([A-H])\d+:[A-H]\d+', range_str)
                        for m in matches:
                            cf_cols_covered.add(m)

        # We expect columns C-H (6 columns) to have CF
        expected_cols = {'C', 'D', 'E', 'F', 'G', 'H'}
        covered = cf_cols_covered & expected_cols
        if len(covered) >= 5:
            print(f"PASS: Component 4 — Conditional formatting <40 red on {len(covered)}/6 percentage cols (0.15 pts)")
            total_score += 0.15
        elif len(covered) >= 1:
            partial = 0.15 * (len(covered) / 6)
            print(f"PARTIAL: Component 4 — CF on {len(covered)}/6 cols ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — No conditional formatting <40 with red fill on percentage columns")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Data validation on Marks sheet (0.15 points)
    # Golden has: whole 0-100 on C2:F21, whole 0-50 on G2:H21
    try:
        dv_list = []
        if ws_marks.data_validations and ws_marks.data_validations.dataValidation:
            dv_list = list(ws_marks.data_validations.dataValidation)

        has_theory_dv = any(
            getattr(dv, 'type', '') in ('whole', 'decimal') and
            str(getattr(dv, 'formula1', '')) == '0' and
            str(getattr(dv, 'formula2', '')) == '100'
            for dv in dv_list
        )
        has_practical_dv = any(
            getattr(dv, 'type', '') in ('whole', 'decimal') and
            str(getattr(dv, 'formula1', '')) == '0' and
            str(getattr(dv, 'formula2', '')) == '50'
            for dv in dv_list
        )

        if has_theory_dv and has_practical_dv:
            print(f"PASS: Component 5 — Data validation for theory (0-100) and practical (0-50) found (0.15 pts)")
            total_score += 0.15
        elif has_theory_dv or has_practical_dv:
            print(f"PARTIAL: Component 5 — Only one validation type found (0.075 pts)")
            total_score += 0.075
        else:
            print(f"FAIL: Component 5 — No appropriate data validation on Marks sheet")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Print layout (landscape orientation) (0.10 points)
    # Initial has orientation=None; golden has orientation=landscape
    try:
        orientation = ws_report.page_setup.orientation
        if orientation and orientation.lower() == 'landscape':
            print(f"PASS: Component 6 — Report sheet has landscape orientation (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 6 — Expected landscape orientation, found: {orientation}")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
