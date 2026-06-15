"""
Initial Setup: Build a dynamic report card system
Task ID: calc_wf_022
Domain: libreoffice_calc

Creates a workbook with three sheets:
  - Subjects: 6 subjects with max marks
  - Marks: 20 students x 6 subjects with raw marks
  - Report: Student names and subject headers only (no formulas)
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_022'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

# ---------- helpers ----------

def launch_gui(command: str, delay_sec: float = 1.0):
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

# ---------- data ----------

SUBJECTS = [
    ("Mathematics",   "Theory",    100),
    ("Physics",       "Theory",    100),
    ("Chemistry",     "Theory",    100),
    ("English",       "Theory",    100),
    ("Computer Lab",  "Practical",  50),
    ("Physics Lab",   "Practical",  50),
]

STUDENTS = [
    "Sarah Chen", "Marcus Johnson", "Priya Patel", "James Wilson",
    "Aisha Mohammed", "Lucas Garcia", "Emma Thompson", "Raj Krishnamurthy",
    "Olivia Brown", "Wei Zhang", "Isabella Martinez", "Ethan Clark",
    "Fatima Al-Rashid", "Noah Williams", "Sophia Lee", "Daniel Kim",
    "Amara Okafor", "Liam O'Brien", "Yuki Tanaka", "Grace Adeyemi",
]

random.seed(42)  # reproducible

def random_marks(max_mark):
    """Generate a realistic spread of marks."""
    # mix of strong, average, and weak students
    marks = []
    for _ in range(len(STUDENTS)):
        band = random.random()
        if band < 0.15:       # weak
            m = random.randint(int(max_mark * 0.15), int(max_mark * 0.38))
        elif band < 0.55:     # average
            m = random.randint(int(max_mark * 0.45), int(max_mark * 0.70))
        elif band < 0.85:     # good
            m = random.randint(int(max_mark * 0.70), int(max_mark * 0.88))
        else:                 # excellent
            m = random.randint(int(max_mark * 0.88), max_mark)
        marks.append(m)
    return marks

# ---------- build workbook ----------

def create_initial():
    wb = openpyxl.Workbook()

    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_font_white = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="000000")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ========== Sheet 1: Subjects ==========
    ws_subj = wb.active
    ws_subj.title = "Subjects"

    subj_headers = ["Subject Code", "Subject Name", "Type", "Max Marks"]
    for c, h in enumerate(subj_headers, 1):
        cell = ws_subj.cell(row=1, column=c, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = box

    for i, (name, stype, maxm) in enumerate(SUBJECTS, 2):
        code = f"SUB{i-1:03d}"
        row_data = [code, name, stype, maxm]
        for c, val in enumerate(row_data, 1):
            cell = ws_subj.cell(row=i, column=c, value=val)
            cell.border = box
            if c in (1, 3, 4):
                cell.alignment = center

    ws_subj.column_dimensions["A"].width = 14
    ws_subj.column_dimensions["B"].width = 18
    ws_subj.column_dimensions["C"].width = 12
    ws_subj.column_dimensions["D"].width = 12

    # ========== Sheet 2: Marks ==========
    ws_marks = wb.create_sheet("Marks")

    marks_headers = ["Roll No", "Student Name"] + [s[0] for s in SUBJECTS]
    for c, h in enumerate(marks_headers, 1):
        cell = ws_marks.cell(row=1, column=c, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = box

    # Generate marks per subject
    all_marks = {}
    for j, (subj_name, stype, maxm) in enumerate(SUBJECTS):
        all_marks[j] = random_marks(maxm)

    for i, student in enumerate(STUDENTS):
        r = i + 2
        ws_marks.cell(row=r, column=1, value=i + 1).border = box
        ws_marks.cell(row=r, column=1).alignment = center
        cell_name = ws_marks.cell(row=r, column=2, value=student)
        cell_name.border = box
        for j in range(len(SUBJECTS)):
            cell = ws_marks.cell(row=r, column=j + 3, value=all_marks[j][i])
            cell.border = box
            cell.alignment = center

    ws_marks.column_dimensions["A"].width = 10
    ws_marks.column_dimensions["B"].width = 22
    for col_letter in ["C", "D", "E", "F", "G", "H"]:
        ws_marks.column_dimensions[col_letter].width = 15

    # ========== Sheet 3: Report ==========
    ws_report = wb.create_sheet("Report")

    # Title row
    ws_report.merge_cells("A1:N1")
    title_cell = ws_report.cell(row=1, column=1, value="Student Report Card - Academic Year 2025-26")
    title_cell.font = Font(name="Calibri", size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Header row (row 3)
    report_headers = [
        "Roll No", "Student Name",
        "Math %", "Physics %", "Chemistry %", "English %",
        "Comp Lab %", "Phys Lab %",
        "Overall %", "GPA", "Grade", "Pass/Fail", "Class Rank",
        "Remarks",
    ]
    for c, h in enumerate(report_headers, 1):
        cell = ws_report.cell(row=3, column=c, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = box

    # Populate Roll No and Student Name only (no formulas)
    for i, student in enumerate(STUDENTS):
        r = i + 4  # data starts row 4
        ws_report.cell(row=r, column=1, value=i + 1).border = box
        ws_report.cell(row=r, column=1).alignment = center
        ws_report.cell(row=r, column=2, value=student).border = box
        # Leave columns C-N empty for the agent task
        for c in range(3, 15):
            ws_report.cell(row=r, column=c).border = box

    ws_report.column_dimensions["A"].width = 10
    ws_report.column_dimensions["B"].width = 22
    for col_letter in ["C", "D", "E", "F", "G", "H", "I"]:
        ws_report.column_dimensions[col_letter].width = 12
    ws_report.column_dimensions["J"].width = 8
    ws_report.column_dimensions["K"].width = 8
    ws_report.column_dimensions["L"].width = 10
    ws_report.column_dimensions["M"].width = 12
    ws_report.column_dimensions["N"].width = 14

    # Freeze panes: freeze header rows
    ws_report.freeze_panes = "A4"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
