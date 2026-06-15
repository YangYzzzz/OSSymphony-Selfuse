"""
Reward Script: Freelance Annual Financial Report with quarterly comparisons,
               tax estimates, and business expense categorization.
Task ID: calc_gpm_095
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25) - Revenue SUM formulas (F5:F9) and Total Revenue row (B10:F10)
  Component 2 (0.20) - Expense SUM formulas (F13:F20) and Total Expenses row (B21:F21)
  Component 3 (0.15) - Profitability formulas (B22:F25 including Net Profit and Profit Margin)
  Component 4 (0.10) - Tax estimate formulas (C28:C31)
  Component 5 (0.15) - Charts (BarChart 'Revenue vs Expenses', PieChart 'Revenue Mix')
  Component 6 (0.15) - Conditional formatting and number formatting
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_095'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists
    if 'AnnualReport' not in wb.sheetnames:
        print("FAIL: Sheet 'AnnualReport' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['AnnualReport']

    # =========================================================================
    # Component 1: Revenue SUM formulas (0.25 points)
    # F5:F9 should have =SUM(Bx:Ex), B10:F10 should have =SUM formulas
    # These are all None in the initial file.
    # =========================================================================
    try:
        c1_score = 0.0
        # Check F5:F9 have SUM formulas
        revenue_sums_ok = 0
        for row in range(5, 10):
            val = ws.cell(row=row, column=6).value  # column F
            if val is not None and isinstance(val, str) and 'SUM' in val.upper():
                revenue_sums_ok += 1
        if revenue_sums_ok == 5:
            c1_score += 0.15
            print(f"PASS: All 5 revenue SUM formulas in F5:F9 present")
        elif revenue_sums_ok > 0:
            partial = 0.15 * (revenue_sums_ok / 5)
            c1_score += partial
            print(f"PARTIAL: {revenue_sums_ok}/5 revenue SUM formulas in F5:F9")
        else:
            print(f"FAIL: No revenue SUM formulas in F5:F9")

        # Check B10:F10 Total Revenue row has SUM formulas
        total_rev_ok = 0
        for col in range(2, 7):  # B through F
            val = ws.cell(row=10, column=col).value
            if val is not None and isinstance(val, str) and 'SUM' in val.upper():
                total_rev_ok += 1
        if total_rev_ok == 5:
            c1_score += 0.10
            print(f"PASS: All 5 Total Revenue SUM formulas in B10:F10 present")
        elif total_rev_ok > 0:
            partial = 0.10 * (total_rev_ok / 5)
            c1_score += partial
            print(f"PARTIAL: {total_rev_ok}/5 Total Revenue formulas in row 10")
        else:
            print(f"FAIL: No Total Revenue formulas in row 10")

        total_score += c1_score
        print(f"Component 1 subtotal: {c1_score:.3f}/0.25")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # =========================================================================
    # Component 2: Expense SUM formulas (0.20 points)
    # F13:F20 should have =SUM, B21:F21 should have =SUM
    # These are all None in initial.
    # =========================================================================
    try:
        c2_score = 0.0
        # Check F13:F20 have SUM formulas
        expense_sums_ok = 0
        for row in range(13, 21):
            val = ws.cell(row=row, column=6).value
            if val is not None and isinstance(val, str) and 'SUM' in val.upper():
                expense_sums_ok += 1
        if expense_sums_ok == 8:
            c2_score += 0.12
            print(f"PASS: All 8 expense SUM formulas in F13:F20 present")
        elif expense_sums_ok > 0:
            partial = 0.12 * (expense_sums_ok / 8)
            c2_score += partial
            print(f"PARTIAL: {expense_sums_ok}/8 expense SUM formulas in F13:F20")
        else:
            print(f"FAIL: No expense SUM formulas in F13:F20")

        # Check B21:F21 Total Expenses row
        total_exp_ok = 0
        for col in range(2, 7):
            val = ws.cell(row=21, column=col).value
            if val is not None and isinstance(val, str) and 'SUM' in val.upper():
                total_exp_ok += 1
        if total_exp_ok == 5:
            c2_score += 0.08
            print(f"PASS: All 5 Total Expenses SUM formulas in B21:F21 present")
        elif total_exp_ok > 0:
            partial = 0.08 * (total_exp_ok / 5)
            c2_score += partial
            print(f"PARTIAL: {total_exp_ok}/5 Total Expenses formulas in row 21")
        else:
            print(f"FAIL: No Total Expenses formulas in row 21")

        total_score += c2_score
        print(f"Component 2 subtotal: {c2_score:.3f}/0.20")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # =========================================================================
    # Component 3: Profitability section formulas (0.15 points)
    # B22:F22 (Gross Revenue = revenue ref), B23:F23 (Total Expenses ref),
    # B24:F24 (Net Profit = Revenue - Expenses), B25:F25 (Profit Margin = ratio)
    # All None in initial.
    # =========================================================================
    try:
        c3_score = 0.0
        # Check that rows 22-25 have formulas in columns B-F
        profitability_formulas = 0
        total_expected = 20  # 4 rows * 5 columns
        for row in range(22, 26):
            for col in range(2, 7):
                val = ws.cell(row=row, column=col).value
                if val is not None and isinstance(val, str) and val.startswith('='):
                    profitability_formulas += 1

        if profitability_formulas >= 18:  # allow minor variance
            c3_score = 0.15
            print(f"PASS: Profitability formulas present ({profitability_formulas}/{total_expected})")
        elif profitability_formulas > 0:
            c3_score = 0.15 * (profitability_formulas / total_expected)
            print(f"PARTIAL: {profitability_formulas}/{total_expected} profitability formulas")
        else:
            print(f"FAIL: No profitability formulas in B22:F25")

        total_score += c3_score
        print(f"Component 3 subtotal: {c3_score:.3f}/0.15")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # =========================================================================
    # Component 4: Tax estimate formulas (0.10 points)
    # C28, C29, C30 should reference net profit * tax rate
    # C31 should be quarterly estimate (sum/4)
    # All None in initial.
    # =========================================================================
    try:
        c4_score = 0.0
        tax_formulas = 0
        for row in [28, 29, 30, 31]:
            val = ws.cell(row=row, column=3).value  # column C
            if val is not None and isinstance(val, str) and val.startswith('='):
                tax_formulas += 1

        if tax_formulas == 4:
            c4_score = 0.10
            print(f"PASS: All 4 tax estimate formulas in C28:C31 present")
        elif tax_formulas > 0:
            c4_score = 0.10 * (tax_formulas / 4)
            print(f"PARTIAL: {tax_formulas}/4 tax formulas present")
        else:
            print(f"FAIL: No tax estimate formulas in C28:C31")

        total_score += c4_score
        print(f"Component 4 subtotal: {c4_score:.3f}/0.10")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # =========================================================================
    # Component 5: Charts (0.15 points)
    # Should have 2 charts: BarChart "Revenue vs Expenses", PieChart "Revenue Mix"
    # Initial has 0 charts.
    # =========================================================================
    try:
        c5_score = 0.0
        charts = ws._charts
        num_charts = len(charts)

        if num_charts >= 2:
            c5_score += 0.05
            print(f"PASS: {num_charts} charts found (expected >= 2)")
        elif num_charts == 1:
            c5_score += 0.025
            print(f"PARTIAL: Only 1 chart found (expected 2)")
        else:
            print(f"FAIL: No charts found")

        # Check for BarChart with 'Revenue vs Expenses' title
        bar_found = False
        pie_found = False
        for ch in charts:
            chart_type = type(ch).__name__
            # Extract title text
            title_text = ''
            try:
                for para in ch.title.tx.rich.paragraphs:
                    for run in para.r:
                        title_text += run.t
            except Exception:
                pass

            if 'Bar' in chart_type and 'revenue' in title_text.lower() and 'expense' in title_text.lower():
                bar_found = True
            if 'Pie' in chart_type and 'revenue' in title_text.lower() and 'mix' in title_text.lower():
                pie_found = True

        if bar_found:
            c5_score += 0.05
            print(f"PASS: BarChart 'Revenue vs Expenses' found")
        else:
            print(f"FAIL: BarChart 'Revenue vs Expenses' not found")

        if pie_found:
            c5_score += 0.05
            print(f"PASS: PieChart 'Revenue Mix' found")
        else:
            print(f"FAIL: PieChart 'Revenue Mix' not found")

        total_score += c5_score
        print(f"Component 5 subtotal: {c5_score:.3f}/0.15")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    # =========================================================================
    # Component 6: Conditional formatting and number formatting (0.15 points)
    # Golden has: color scale on B5:E9, conditional rules on B25:F25,
    # data bars on F5:F9 and F13:F20, $#,##0 number format on currency cells,
    # 0.0% on profit margin cells.
    # Initial has NO conditional formatting at all.
    # =========================================================================
    try:
        c6_score = 0.0

        # Check conditional formatting exists (initial has 0 rules)
        cf_ranges = []
        for cf in ws.conditional_formatting:
            cf_ranges.append(str(cf))

        # Check for color scale on revenue data (B5:E9)
        has_color_scale = False
        for cf in ws.conditional_formatting:
            for rule in cf.rules:
                if rule.type == 'colorScale':
                    has_color_scale = True
                    break

        if has_color_scale:
            c6_score += 0.03
            print(f"PASS: Color scale conditional formatting found")
        else:
            print(f"FAIL: No color scale conditional formatting")

        # Check for profit margin conditional formatting (cellIs rules)
        has_profit_cf = False
        for cf in ws.conditional_formatting:
            for rule in cf.rules:
                if rule.type == 'cellIs':
                    has_profit_cf = True
                    break

        if has_profit_cf:
            c6_score += 0.04
            print(f"PASS: Profit margin conditional formatting found")
        else:
            print(f"FAIL: No profit margin conditional formatting")

        # Check for data bars
        has_data_bars = False
        for cf in ws.conditional_formatting:
            for rule in cf.rules:
                if rule.type == 'dataBar':
                    has_data_bars = True
                    break

        if has_data_bars:
            c6_score += 0.03
            print(f"PASS: Data bar conditional formatting found")
        else:
            print(f"FAIL: No data bar conditional formatting")

        # Check number format on revenue total cells (should be $#,##0)
        currency_format_ok = 0
        for coord in ['F5', 'F6', 'F7', 'F8', 'F9']:
            cell = ws[coord]
            if cell.number_format is not None and '$' in str(cell.number_format):
                currency_format_ok += 1

        if currency_format_ok >= 3:
            c6_score += 0.025
            print(f"PASS: Currency number format on revenue totals ({currency_format_ok}/5)")
        else:
            print(f"FAIL: Currency number format missing on revenue totals ({currency_format_ok}/5)")

        # Check percentage format on profit margin row (B25:F25)
        # AND that formulas exist there (format alone is a precondition in initial)
        pct_format_ok = 0
        for col in range(2, 7):
            cell = ws.cell(row=25, column=col)
            has_formula = cell.value is not None and isinstance(cell.value, str) and cell.value.startswith('=')
            has_pct = cell.number_format is not None and '%' in str(cell.number_format)
            if has_formula and has_pct:
                pct_format_ok += 1

        if pct_format_ok >= 3:
            c6_score += 0.025
            print(f"PASS: Percentage format + formulas on profit margin row ({pct_format_ok}/5)")
        else:
            print(f"FAIL: Percentage format + formulas missing on profit margin ({pct_format_ok}/5)")

        total_score += c6_score
        print(f"Component 6 subtotal: {c6_score:.3f}/0.15")
    except Exception as e:
        print(f"ERROR: Component 6 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score:.3f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
