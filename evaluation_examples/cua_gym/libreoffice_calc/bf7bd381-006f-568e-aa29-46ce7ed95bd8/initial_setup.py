"""
Initial Setup: Freelance Annual Financial Report with quarterly data
Task ID: calc_gpm_095
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_095'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'AnnualReport'

    # ---- Styles ----
    thick = Side(style='thick', color='000000')
    thin = Side(style='thin', color='000000')
    medium = Side(style='medium', color='000000')
    thick_border = Border(left=thick, right=thick, top=thick, bottom=thick)
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    dollar_fmt = '$#,##0'
    pct_fmt = '0.0%'

    # ---- Title Row (A1:J1) ----
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = 'Freelance Annual Financial Report - 2025'
    title_cell.font = Font(size=16, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # ---- Subtitle Row (A2:J2) ----
    ws.merge_cells('A2:J2')
    sub_cell = ws['A2']
    sub_cell.value = 'Prepared for: Alex Torres | Business: Torres Design Studio | EIN: XX-XXXXXXX'
    sub_cell.font = Font(size=10, italic=True)
    sub_cell.fill = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')
    sub_cell.alignment = Alignment(horizontal='center', vertical='center')

    # ---- Column widths ----
    ws.column_dimensions['A'].width = 22
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col_letter].width = 14
    for col_letter in ['H', 'I', 'J']:
        ws.column_dimensions[col_letter].width = 14

    # ========================================================
    # Section 1: Quarterly Revenue (A4:F9)
    # ========================================================
    rev_headers = ['Revenue Source', 'Q1', 'Q2', 'Q3', 'Q4', 'Total']
    for c, h in enumerate(rev_headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    revenue_data = [
        ['Client Projects',    18500, 22300, 25100, 27800],
        ['Retainer Contracts',  8400,  8400,  9200,  9200],
        ['Workshops',           3200,  4500,  2800,  5100],
        ['Digital Products',    1800,  2100,  2600,  3400],
        ['Consulting',          5600,  6200,  7100,  8300],
    ]
    for r, row_data in enumerate(revenue_data, 5):
        ws.cell(row=r, column=1, value=row_data[0]).border = thin_border
        for c, val in enumerate(row_data[1:], 2):
            cell = ws.cell(row=r, column=c, value=val)
            cell.number_format = dollar_fmt
            cell.border = thin_border
        # F column (Total) left EMPTY for the task
        ws.cell(row=r, column=6).border = thin_border

    # Row 10: Total Revenue row - label only, no formulas
    ws.cell(row=10, column=1, value='Total Revenue').font = Font(bold=True)
    ws.cell(row=10, column=1).border = thin_border
    for c in range(2, 7):
        cell = ws.cell(row=10, column=c)
        cell.border = thin_border
        cell.number_format = dollar_fmt

    # ========================================================
    # Section 2: Expenses (A12:F20)
    # ========================================================
    exp_headers = ['Expense Category', 'Q1', 'Q2', 'Q3', 'Q4', 'Total']
    for c, h in enumerate(exp_headers, 1):
        cell = ws.cell(row=12, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    expense_data = [
        ['Office Rent',        2400, 2400, 2400, 2400],
        ['Software/Tools',      350,  350,  420,  420],
        ['Hardware',            800,    0, 1200,    0],
        ['Marketing',           600,  750,  900, 1100],
        ['Insurance',           450,  450,  450,  450],
        ['Travel',              800, 1200,  600, 1500],
        ['Professional Dev',    500,  300,  700,  400],
        ['Misc',                200,  250,  180,  320],
    ]
    for r, row_data in enumerate(expense_data, 13):
        ws.cell(row=r, column=1, value=row_data[0]).border = thin_border
        for c, val in enumerate(row_data[1:], 2):
            cell = ws.cell(row=r, column=c, value=val)
            cell.number_format = dollar_fmt
            cell.border = thin_border
        # F column (Total) left EMPTY
        ws.cell(row=r, column=6).border = thin_border

    # Row 21: Subtotal row placeholder (no formulas)
    ws.cell(row=21, column=1, value='Total Expenses').font = Font(bold=True)
    ws.cell(row=21, column=1).border = thin_border
    for c in range(2, 7):
        cell = ws.cell(row=21, column=c)
        cell.border = thin_border
        cell.number_format = dollar_fmt

    # ========================================================
    # Section 3: Profitability (A22:F25)
    # ========================================================
    prof_labels = ['Gross Revenue', 'Total Expenses', 'Net Profit', 'Profit Margin']
    for r, label in enumerate(prof_labels, 22):
        cell = ws.cell(row=r, column=1, value=label)
        cell.font = Font(bold=True)
        cell.border = thin_border
        for c in range(2, 7):
            ws.cell(row=r, column=c).border = thin_border
    # Profit Margin row format
    for c in range(2, 7):
        ws.cell(row=25, column=c).number_format = pct_fmt
        ws.cell(row=22, column=c).number_format = dollar_fmt
        ws.cell(row=23, column=c).number_format = dollar_fmt
        ws.cell(row=24, column=c).number_format = dollar_fmt

    # ========================================================
    # Section 4: Tax Estimates (A27:C31)
    # ========================================================
    ws.cell(row=27, column=1, value='Tax Category').font = Font(bold=True, color='FFFFFF')
    ws.cell(row=27, column=1).fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
    ws.cell(row=27, column=1).border = thin_border

    ws.cell(row=27, column=2, value='Rate').font = Font(bold=True, color='FFFFFF')
    ws.cell(row=27, column=2).fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
    ws.cell(row=27, column=2).border = thin_border

    ws.cell(row=27, column=3, value='Estimated Amount').font = Font(bold=True, color='FFFFFF')
    ws.cell(row=27, column=3).fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
    ws.cell(row=27, column=3).border = thin_border

    tax_data = [
        ['Self-Employment Tax', '15.3%'],
        ['Estimated Federal Tax', '24%'],
        ['State Tax', '5%'],
        ['Total Quarterly Estimate', ''],
    ]
    for r, row_data in enumerate(tax_data, 28):
        ws.cell(row=r, column=1, value=row_data[0]).border = thin_border
        ws.cell(row=r, column=2, value=row_data[1]).border = thin_border
        cell = ws.cell(row=r, column=3)
        cell.border = thin_border
        cell.number_format = dollar_fmt

    # ---- Apply thick borders around each section ----
    # Section 1: A4:F10
    for r in range(4, 11):
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            left = thick if c == 1 else thin
            right = thick if c == 6 else thin
            top = thick if r == 4 else thin
            bottom = thick if r == 10 else thin
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    # Section 2: A12:F21
    for r in range(12, 22):
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            left = thick if c == 1 else thin
            right = thick if c == 6 else thin
            top = thick if r == 12 else thin
            bottom = thick if r == 21 else thin
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    # Section 3: A22:F25
    for r in range(22, 26):
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            left = thick if c == 1 else thin
            right = thick if c == 6 else thin
            top = thick if r == 22 else thin
            bottom = thick if r == 25 else thin
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    # Section 4: A27:C31
    for r in range(27, 32):
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            left = thick if c == 1 else thin
            right = thick if c == 3 else thin
            top = thick if r == 27 else thin
            bottom = thick if r == 31 else thin
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
