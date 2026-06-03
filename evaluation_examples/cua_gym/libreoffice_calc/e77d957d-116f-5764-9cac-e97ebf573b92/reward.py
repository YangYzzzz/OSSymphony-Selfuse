"""
Reward Script: Add 'channels_to_follow' sheet with top-30 YouTube channels not already followed
Task ID: osworld_multi_apps_misc_020
Domain: libreoffice_calc
Scoring:
  - Component 1: 'channels_to_follow' sheet exists (0.20 pts)
  - Component 2: Correct headers (Rank, Channel Name, Subscribers (M), Category) (0.20 pts)
  - Component 3: All 15 expected channel ranks are present (0.40 pts)
  - Component 4: Rows sorted by Rank ascending (0.20 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'youtube_channels'

# Expected channels_to_follow data (top-30 from SocialBlade, created before 2022, not in my_channels)
# These are the 15 channels in the golden_env channels_to_follow sheet, identified by rank.
EXPECTED_RANKS = {4, 6, 9, 11, 13, 14, 16, 18, 19, 21, 23, 24, 26, 28, 29}
EXPECTED_HEADERS = ('Rank', 'Channel Name', 'Subscribers (M)', 'Category')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook — precondition gate
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: 'channels_to_follow' sheet exists (0.20 pts)
    # This FAILS on initial (only 'my_channels' sheet) and PASSES on golden
    try:
        if 'channels_to_follow' in wb.sheetnames:
            print("PASS: Component 1 — 'channels_to_follow' sheet exists (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 — 'channels_to_follow' sheet not found. Sheets: {wb.sheetnames}")
            # If the sheet doesn't exist, remaining components cannot be checked
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws = wb['channels_to_follow']

    # Component 2: Correct headers in row 1 (0.20 pts)
    # This FAILS on initial (sheet doesn't exist) and PASSES on golden
    try:
        row1 = tuple(ws.cell(row=1, column=c).value for c in range(1, 5))
        if row1 == EXPECTED_HEADERS:
            print(f"PASS: Component 2 — Correct headers: {row1} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 — Expected headers {EXPECTED_HEADERS}, found: {row1}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: All 15 expected channel ranks present (0.40 pts)
    # This FAILS on initial (sheet doesn't exist) and PASSES on golden
    try:
        actual_ranks = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                try:
                    actual_ranks.add(int(row[0]))
                except (TypeError, ValueError):
                    pass

        missing_ranks = EXPECTED_RANKS - actual_ranks
        extra_ranks = actual_ranks - EXPECTED_RANKS

        if not missing_ranks and not extra_ranks:
            print(f"PASS: Component 3 — All 15 expected channel ranks present: {sorted(actual_ranks)} (0.40 pts)")
            total_score += 0.40
        else:
            if missing_ranks:
                print(f"FAIL: Component 3 — Missing ranks from channels_to_follow: {sorted(missing_ranks)}")
            if extra_ranks:
                print(f"FAIL: Component 3 — Unexpected ranks in channels_to_follow: {sorted(extra_ranks)}")
            # Partial credit: 0.40 * (correct / 15)
            correct_count = len(EXPECTED_RANKS & actual_ranks)
            partial = round(0.40 * correct_count / 15, 4)
            print(f"PARTIAL: Component 3 — {correct_count}/15 correct ranks found, awarding {partial} pts")
            total_score += partial
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Rows sorted by Rank ascending (0.20 pts)
    # This FAILS on initial (sheet doesn't exist) and PASSES on golden
    try:
        ranks_in_order = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                try:
                    ranks_in_order.append(int(row[0]))
                except (TypeError, ValueError):
                    pass

        if ranks_in_order and ranks_in_order == sorted(ranks_in_order):
            print(f"PASS: Component 4 — Rows sorted by Rank ascending: {ranks_in_order} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 4 — Rows not sorted by Rank ascending. Found order: {ranks_in_order}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in given env
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
