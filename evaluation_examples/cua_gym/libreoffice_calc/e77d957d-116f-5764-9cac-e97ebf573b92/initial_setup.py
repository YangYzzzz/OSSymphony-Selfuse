"""
Initial Setup: YouTube channels followed list in LibreOffice Calc + Chrome at SocialBlade
Task ID: osworld_multi_apps_misc_020
Domain: libreoffice_calc + chrome (multi-app)

Creates youtube_channels.xlsx with the user's current list of followed YouTube channels,
then opens it in LibreOffice Calc and opens Chrome at the SocialBlade top YouTube channels page.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_misc_020'
OUTPUT = f'{WORKDIR}/youtube_channels.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: My Channels (user's followed list) ---
    ws = wb.active
    ws.title = 'my_channels'

    # Column headers
    headers = ['Rank', 'Channel Name', 'Subscribers (M)', 'Category']
    header_font = Font(bold=True, name='Calibri', size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    # User's currently followed channels (a subset of the top 30 channels from SocialBlade)
    # These are channels the user ALREADY follows — they should NOT appear in channels_to_follow
    # Channels are identified by their rank in the SocialBlade top 30 list
    followed_channels = [
        # Rank, Channel Name,              Subscribers (M), Category
        (1,  'T-Series',                    270.0,  'Music'),
        (2,  'MrBeast',                     252.0,  'Entertainment'),
        (3,  'Cocomelon - Nursery Rhymes',  180.0,  'Education'),
        (5,  'Kids Diana Show',             124.0,  'Entertainment'),
        (7,  'Like Nastya',                 120.0,  'Entertainment'),
        (8,  'Vlad and Niki',               109.0,  'Entertainment'),
        (10, 'Zee Music Company',           107.0,  'Music'),
        (12, 'BLACKPINK',                    95.0,  'Music'),
        (15, 'Pinkfong Baby Shark',          79.0,  'Education'),
        (17, 'Justin Bieber',                73.0,  'Music'),
        (20, 'Dua Lipa',                     52.0,  'Music'),
        (22, 'Ed Sheeran',                   51.0,  'Music'),
        (25, 'Taylor Swift',                 49.0,  'Music'),
        (27, 'Eminem',                       48.0,  'Music'),
        (30, 'Marshmello',                   46.0,  'Music'),
    ]

    for r, row_data in enumerate(followed_channels, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Freeze the header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open LibreOffice Calc with the spreadsheet
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=3.0)

    # Open Chrome at the SocialBlade top YouTube channels page
    launch_gui('google-chrome "https://socialblade.com/youtube/top"', delay_sec=3.0)

    print('GUI_READY: launched LibreOffice Calc and Chrome with DISPLAY=:0')


create_initial()
