"""
Initial Setup: Add diagonal cross-hatch border style to C5:E9 on Design sheet
Task ID: calc_gg3_028
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_028'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Design (layout planning tool) ---
    ws = wb.active
    ws.title = 'Design'

    # Title row
    ws.merge_cells('A1:H1')
    ws['A1'] = 'Website Layout Mockup — Homepage Redesign 2025'
    ws['A1'].font = Font(name='Arial', size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Subtitle
    ws.merge_cells('A2:H2')
    ws['A2'] = 'Grid zones: each cell represents a 120px block'
    ws['A2'].font = Font(name='Arial', size=10, italic=True, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')

    # Column headers (grid labels)
    col_labels = ['', 'Col A', 'Col B', 'Col C', 'Col D', 'Col E', 'Col F', 'Col G']
    for c, label in enumerate(col_labels, 1):
        cell = ws.cell(row=3, column=c, value=label)
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='center')
        if c > 1:
            cell.fill = PatternFill(start_color='FFD9E2F3', end_color='FFD9E2F3', fill_type='solid')

    # Row labels and zone descriptions
    row_labels = {
        4: 'Row 1',
        5: 'Row 2',
        6: 'Row 3',
        7: 'Row 4',
        8: 'Row 5',
        9: 'Row 6',
        10: 'Row 7',
        11: 'Row 8',
        12: 'Row 9',
    }
    for r, label in row_labels.items():
        cell = ws.cell(row=r, column=1, value=label)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color='FFD9E2F3', end_color='FFD9E2F3', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Fill in zone descriptions (layout planning content)
    # Row 4 - navigation bar zone
    zone_data = {
        (4, 2): 'Nav Logo',
        (4, 3): 'Nav Menu',
        (4, 4): 'Nav Menu',
        (4, 5): 'Nav Search',
        (4, 6): 'Nav Login',
        (4, 7): 'Nav Cart',
        (4, 8): 'Nav Profile',
        # Row 5-9: C5:E9 is the "unavailable" blocked zone (no borders/fills here!)
        (5, 2): 'Sidebar',
        (5, 6): 'Hero Image',
        (5, 7): 'Hero Image',
        (5, 8): 'Hero Image',
        (6, 2): 'Sidebar',
        (6, 6): 'Feature 1',
        (6, 7): 'Feature 2',
        (6, 8): 'Feature 3',
        (7, 2): 'Sidebar',
        (7, 6): 'Pricing A',
        (7, 7): 'Pricing B',
        (7, 8): 'Pricing C',
        (8, 2): 'Sidebar',
        (8, 6): 'CTA Button',
        (8, 7): 'CTA Button',
        (8, 8): 'Newsletter',
        (9, 2): 'Sidebar',
        (9, 6): 'Footer Left',
        (9, 7): 'Footer Mid',
        (9, 8): 'Footer Right',
        # Row 10-12: additional content
        (10, 2): 'Ad Space',
        (10, 3): 'Blog Post 1',
        (10, 4): 'Blog Post 2',
        (10, 5): 'Blog Post 3',
        (10, 6): 'Blog Post 4',
        (10, 7): 'Blog Post 5',
        (10, 8): 'Social Feed',
        (11, 2): 'Ad Space',
        (11, 3): 'Reviews',
        (11, 4): 'Reviews',
        (11, 5): 'FAQ Section',
        (11, 6): 'FAQ Section',
        (11, 7): 'Contact Us',
        (11, 8): 'Map Widget',
        (12, 2): 'Copyright',
        (12, 3): 'Privacy',
        (12, 4): 'Terms',
        (12, 5): 'Sitemap',
        (12, 6): 'Accessibility',
        (12, 7): 'Language',
        (12, 8): 'Back to Top',
    }

    # C5:E9 = columns 3-5, rows 5-9: content but NO borders/fills
    # These are the "unavailable" zones that the task will mark
    unavailable_labels = {
        (5, 3): 'Reserved',
        (5, 4): 'Reserved',
        (5, 5): 'Reserved',
        (6, 3): 'Blocked',
        (6, 4): 'Blocked',
        (6, 5): 'Blocked',
        (7, 3): 'Pending',
        (7, 4): 'Pending',
        (7, 5): 'Pending',
        (8, 3): 'TBD',
        (8, 4): 'TBD',
        (8, 5): 'TBD',
        (9, 3): 'Unused',
        (9, 4): 'Unused',
        (9, 5): 'Unused',
    }

    for (r, c), val in zone_data.items():
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(size=9)

    for (r, c), val in unavailable_labels.items():
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(size=9, color='999999')
        # NO borders, NO fills on C5:E9

    # Set column widths
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col_letter].width = 14

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

    # --- Sheet 2: Legend ---
    ws2 = wb.create_sheet('Legend')
    ws2['A1'] = 'Zone Type'
    ws2['B1'] = 'Description'
    ws2['C1'] = 'Status'
    ws2['A1'].font = Font(bold=True)
    ws2['B1'].font = Font(bold=True)
    ws2['C1'].font = Font(bold=True)

    legend_data = [
        ['Navigation', 'Top navigation bar elements', 'Finalized'],
        ['Sidebar', 'Left sidebar navigation and widgets', 'In Review'],
        ['Content Area', 'Main content blocks for features/pricing', 'Active'],
        ['Hero Section', 'Large hero image/banner area', 'Active'],
        ['Footer', 'Bottom page footer elements', 'Draft'],
        ['Blocked Zone', 'Areas reserved for future development', 'Unavailable'],
        ['Ad Space', 'Advertising placement zones', 'Pending Approval'],
        ['Blog Section', 'Blog post preview cards', 'In Development'],
        ['Contact/FAQ', 'Support and information sections', 'Planned'],
    ]
    for r, row_data in enumerate(legend_data, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 40
    ws2.column_dimensions['C'].width = 18

    # --- Sheet 3: Notes ---
    ws3 = wb.create_sheet('Notes')
    ws3['A1'] = 'Design Notes'
    ws3['A1'].font = Font(bold=True, size=12)
    ws3['A2'] = 'Last updated: 2025-03-28'
    ws3['A3'] = 'Lead designer: Elena Rodriguez'
    ws3['A4'] = 'Project: Website Redesign Q2 2025'
    ws3['A6'] = 'The Design sheet uses a grid layout where each cell represents a 120px block.'
    ws3['A7'] = 'Cells marked Reserved/Blocked/Pending/TBD/Unused in the C5:E9 range'
    ws3['A8'] = 'should be visually marked as unavailable using cell formatting.'
    ws3['A9'] = 'The designer needs to apply a cross-hatch diagonal border pattern'
    ws3['A10'] = 'and a thick blue outer border to clearly distinguish these blocked zones.'
    ws3.column_dimensions['A'].width = 70

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
