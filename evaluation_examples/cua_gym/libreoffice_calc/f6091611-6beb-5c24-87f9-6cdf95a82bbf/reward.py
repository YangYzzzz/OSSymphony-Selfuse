"""
Reward Script: Diagonal cross-hatch borders + thick blue outer border on C5:E9
Task ID: calc_gg3_028
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Diagonal borders (both diagonalDown + diagonalUp) in all 15 cells of C5:E9
  Component 2 (0.4): Thick blue outer border around the perimeter of C5:E9
  Component 3 (0.2): No spurious border changes on cells outside C5:E9
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_028'


def persist_app_state(domain: str):
    """Try to save any unsaved LibreOffice state via Ctrl+S."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that 'Design' sheet exists
    if 'Design' not in wb.sheetnames:
        print("CRITICAL: 'Design' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Design']

    # Target range: C5:E9 (row 5-9, col 3-5)
    TARGET_MIN_ROW = 5
    TARGET_MAX_ROW = 9
    TARGET_MIN_COL = 3  # C
    TARGET_MAX_COL = 5  # E

    # ---------------------------------------------------------------
    # Component 1: Diagonal borders in ALL cells of C5:E9 (0.4 pts)
    # Both diagonalDown and diagonalUp should be True with a non-None style
    # ---------------------------------------------------------------
    try:
        total_cells = (TARGET_MAX_ROW - TARGET_MIN_ROW + 1) * (TARGET_MAX_COL - TARGET_MIN_COL + 1)
        diag_pass = 0
        for r in range(TARGET_MIN_ROW, TARGET_MAX_ROW + 1):
            for c in range(TARGET_MIN_COL, TARGET_MAX_COL + 1):
                cell = ws.cell(row=r, column=c)
                b = cell.border
                has_diag_style = b.diagonal.style is not None
                has_diag_down = b.diagonalDown is True
                has_diag_up = b.diagonalUp is True
                if has_diag_style and has_diag_down and has_diag_up:
                    diag_pass += 1
                else:
                    coord = cell.coordinate
                    print(f"FAIL: Component 1 — {coord} missing diagonal cross-hatch "
                          f"(style={b.diagonal.style}, down={b.diagonalDown}, up={b.diagonalUp})")

        if diag_pass == total_cells:
            print(f"PASS: Component 1 — All {total_cells} cells in C5:E9 have diagonal cross-hatch (0.4 pts)")
            total_score += 0.4
        elif diag_pass > 0:
            partial = round(0.4 * diag_pass / total_cells, 2)
            print(f"PARTIAL: Component 1 — {diag_pass}/{total_cells} cells have diagonal cross-hatch ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No cells in C5:E9 have diagonal cross-hatch borders")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ---------------------------------------------------------------
    # Component 2: Thick blue outer border around C5:E9 (0.4 pts)
    # Left edge: cells C5-C9 must have thick blue left border
    # Right edge: cells E5-E9 must have thick blue right border
    # Top edge: cells C5-E5 must have thick blue top border
    # Bottom edge: cells C9-E9 must have thick blue bottom border
    # ---------------------------------------------------------------
    try:
        outer_checks = 0
        outer_pass = 0

        def is_thick_blue(side):
            """Check if a border side is thick and blue."""
            if side.style != 'thick':
                return False
            if side.color is None:
                return False
            rgb = getattr(side.color, 'rgb', None)
            if rgb is None:
                return False
            # Blue = FF0000FF (ARGB) or 0000FF (RGB portion)
            rgb_str = str(rgb).upper()
            # Accept any ARGB where R=00, G=00, B=FF (pure blue)
            # Common: FF0000FF, 000000FF
            return rgb_str.endswith('0000FF')

        # Left edge: C5 through C9
        for r in range(TARGET_MIN_ROW, TARGET_MAX_ROW + 1):
            outer_checks += 1
            cell = ws.cell(row=r, column=TARGET_MIN_COL)
            if is_thick_blue(cell.border.left):
                outer_pass += 1
            else:
                print(f"FAIL: Component 2 — {cell.coordinate} left border not thick blue "
                      f"(style={cell.border.left.style}, color={getattr(cell.border.left.color, 'rgb', None)})")

        # Right edge: E5 through E9
        for r in range(TARGET_MIN_ROW, TARGET_MAX_ROW + 1):
            outer_checks += 1
            cell = ws.cell(row=r, column=TARGET_MAX_COL)
            if is_thick_blue(cell.border.right):
                outer_pass += 1
            else:
                print(f"FAIL: Component 2 — {cell.coordinate} right border not thick blue "
                      f"(style={cell.border.right.style}, color={getattr(cell.border.right.color, 'rgb', None)})")

        # Top edge: C5 through E5
        for c in range(TARGET_MIN_COL, TARGET_MAX_COL + 1):
            outer_checks += 1
            cell = ws.cell(row=TARGET_MIN_ROW, column=c)
            if is_thick_blue(cell.border.top):
                outer_pass += 1
            else:
                print(f"FAIL: Component 2 — {cell.coordinate} top border not thick blue "
                      f"(style={cell.border.top.style}, color={getattr(cell.border.top.color, 'rgb', None)})")

        # Bottom edge: C9 through E9
        for c in range(TARGET_MIN_COL, TARGET_MAX_COL + 1):
            outer_checks += 1
            cell = ws.cell(row=TARGET_MAX_ROW, column=c)
            if is_thick_blue(cell.border.bottom):
                outer_pass += 1
            else:
                print(f"FAIL: Component 2 — {cell.coordinate} bottom border not thick blue "
                      f"(style={cell.border.bottom.style}, color={getattr(cell.border.bottom.color, 'rgb', None)})")

        if outer_pass == outer_checks:
            print(f"PASS: Component 2 — All {outer_checks} outer border edges are thick blue (0.4 pts)")
            total_score += 0.4
        elif outer_pass > 0:
            partial = round(0.4 * outer_pass / outer_checks, 2)
            print(f"PARTIAL: Component 2 — {outer_pass}/{outer_checks} outer edges are thick blue ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No outer border edges are thick blue")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---------------------------------------------------------------
    # Component 3: Borders are correctly scoped — diagonal borders
    # exist in C5:E9 AND no diagonal borders exist outside C5:E9 (0.2 pts)
    # This is a compound check anchored to the task change: the first
    # sub-condition (diagonals present inside range) FAILS on initial_env.
    # ---------------------------------------------------------------
    try:
        # Sub-condition A: At least one cell in C5:E9 has diagonal borders
        has_diag_inside = False
        for r in range(TARGET_MIN_ROW, TARGET_MAX_ROW + 1):
            for c in range(TARGET_MIN_COL, TARGET_MAX_COL + 1):
                cell = ws.cell(row=r, column=c)
                if cell.border.diagonal.style is not None:
                    has_diag_inside = True
                    break
            if has_diag_inside:
                break

        # Sub-condition B: No cells outside C5:E9 have diagonal borders
        diag_outside = 0
        for r in range(1, 13):
            for c in range(1, 9):
                if TARGET_MIN_ROW <= r <= TARGET_MAX_ROW and TARGET_MIN_COL <= c <= TARGET_MAX_COL:
                    continue
                cell = ws.cell(row=r, column=c)
                if cell.border.diagonal.style is not None:
                    diag_outside += 1
                    print(f"FAIL: Component 3 — {cell.coordinate} has unexpected diagonal border")

        if has_diag_inside and diag_outside == 0:
            print(f"PASS: Component 3 — Diagonal borders correctly scoped to C5:E9 only (0.2 pts)")
            total_score += 0.2
        elif not has_diag_inside:
            print(f"FAIL: Component 3 — No diagonal borders found inside C5:E9 (precondition for scoping check)")
        else:
            print(f"FAIL: Component 3 — {diag_outside} cells outside C5:E9 have unexpected diagonal borders")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
