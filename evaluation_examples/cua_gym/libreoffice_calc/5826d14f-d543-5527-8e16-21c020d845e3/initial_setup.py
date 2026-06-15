"""
Initial Setup: SalesReps spreadsheet with 90 rows of data
Task ID: calc_dop_filter_multi_074
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_dop_filter_multi_074'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'SalesReps'

    # Headers
    headers = ['Rep ID', 'Name', 'Region', 'Status', 'Revenue', 'Quota', 'Attainment']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # AutoFilter - enabled but no filter conditions applied
    ws.auto_filter.ref = 'A1:G91'

    # Freeze the header row
    ws.freeze_panes = 'A2'

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14

    # Define data rows
    # Exactly 13 rows must match: Status=Active AND Region=North AND Revenue>25000
    # These will be mixed throughout the dataset

    data = [
        # (Rep ID, Name, Region, Status, Revenue, Quota, Attainment)
        # --- Rows that match all 3 criteria: Active + North + Revenue > 25000 ---
        ('R001', 'Nathan Caldwell',    'North', 'Active',   47200, 45000, 104.9),
        ('R002', 'Priya Sharma',       'North', 'Active',   63500, 60000, 105.8),
        ('R003', 'Derek Whitman',      'North', 'Active',   38900, 35000, 111.1),
        ('R004', 'Chloe Beaumont',     'North', 'Active',   52400, 50000, 104.8),
        ('R005', 'Elijah Monroe',      'North', 'Active',   29300, 28000, 104.6),
        ('R006', 'Isabelle Laurent',   'North', 'Active',   71800, 70000, 102.6),
        ('R007', 'Marcus Osei',        'North', 'Active',   44100, 42000, 105.0),
        ('R008', 'Tara Nishida',       'North', 'Active',   33600, 32000, 105.0),
        ('R009', 'Brendan Kowalski',   'North', 'Active',   58700, 55000, 106.7),
        ('R010', 'Simone Delacroix',   'North', 'Active',   27800, 27000, 103.0),
        ('R011', 'Ryan Petersen',      'North', 'Active',   83200, 80000, 104.0),
        ('R012', 'Anika Johansson',    'North', 'Active',   39600, 38000, 104.2),
        ('R013', 'Quinton Hargrove',   'North', 'Active',   65400, 62000, 105.5),
        # --- North + Active but Revenue <= 25000 (does NOT fully match) ---
        ('R014', 'Lena Fitzgerald',    'North', 'Active',   21500, 25000,  86.0),
        ('R015', 'Omar Castillo',      'North', 'Active',   18900, 22000,  85.9),
        # --- North + Revenue > 25000 but Inactive ---
        ('R016', 'Heather Bloom',      'North', 'Inactive', 48300, 45000, 107.3),
        ('R017', 'Jason Drummond',     'North', 'Inactive', 36700, 35000, 104.9),
        ('R018', 'Sofia Vargas',       'North', 'On Leave', 42000, 40000, 105.0),
        # --- South + Active + Revenue > 25000 ---
        ('R019', 'Gerald Tompkins',    'South', 'Active',   55200, 52000, 106.2),
        ('R020', 'Alicia Moreno',      'South', 'Active',   31400, 30000, 104.7),
        ('R021', 'Darius King',        'South', 'Active',   47800, 45000, 106.2),
        ('R022', 'Nora Sullivan',      'South', 'Active',   29100, 28000, 103.9),
        ('R023', 'Theo Baines',        'South', 'Active',   63900, 60000, 106.5),
        ('R024', 'Valentina Cruz',     'South', 'Active',   38500, 37000, 104.1),
        # --- South + Inactive ---
        ('R025', 'Miles Jefferson',    'South', 'Inactive', 44500, 43000, 103.5),
        ('R026', 'Carmen Navarro',     'South', 'Inactive', 27600, 27000, 102.2),
        ('R027', 'Leon Graves',        'South', 'On Leave', 19200, 22000,  87.3),
        # --- East + Active + Revenue > 25000 ---
        ('R028', 'Adriana Patel',      'East',  'Active',   72300, 70000, 103.3),
        ('R029', 'Flynn McCarthy',     'East',  'Active',   33800, 32000, 105.6),
        ('R030', 'Gwendolyn Park',     'East',  'Active',   26500, 26000, 101.9),
        ('R031', 'Hassan Yusuf',       'East',  'Active',   58100, 55000, 105.6),
        ('R032', 'Iris Nakamura',      'East',  'Active',   44700, 43000, 104.0),
        ('R033', 'Jordan Blackwell',   'East',  'Active',   37200, 36000, 103.3),
        # --- East + Inactive/On Leave ---
        ('R034', 'Keisha Thornton',    'East',  'Inactive', 61200, 60000, 102.0),
        ('R035', 'Lorenzo Ferreira',   'East',  'Inactive', 29800, 29000, 102.8),
        ('R036', 'Megan Forsythe',     'East',  'On Leave', 53400, 52000, 102.7),
        # --- West + Active + Revenue > 25000 ---
        ('R037', 'Nicolas Rousseau',   'West',  'Active',   41900, 40000, 104.8),
        ('R038', 'Olivia Stephenson',  'West',  'Active',   68200, 65000, 104.9),
        ('R039', 'Patrick Oluwole',    'West',  'Active',   56700, 54000, 105.0),
        ('R040', 'Queenie Hammond',    'West',  'Active',   30100, 30000, 100.3),
        ('R041', 'Rebecca Tran',       'West',  'Active',   49500, 48000, 103.1),
        ('R042', 'Samuel Brinkley',    'West',  'Active',   25800, 25500, 101.2),
        # --- West + Inactive ---
        ('R043', 'Tanya Goldstein',    'West',  'Inactive', 37600, 36000, 104.4),
        ('R044', 'Ulysses Adeyemi',    'West',  'Inactive', 22400, 23000,  97.4),
        ('R045', 'Victoria Larson',    'West',  'On Leave', 44800, 43000, 104.2),
        # --- North + Active but low revenue ---
        ('R046', 'Walter Hennessy',    'North', 'Active',   14500, 18000,  80.6),
        ('R047', 'Xena Korhonen',      'North', 'Active',    9800, 12000,  81.7),
        # --- More mixed data ---
        ('R048', 'Yolanda Mukherjee',  'South', 'Active',   24800, 25000,  99.2),
        ('R049', 'Zachary Henriksen',  'East',  'Active',   11200, 14000,  80.0),
        ('R050', 'Abigail Torres',     'West',  'Active',   57300, 55000, 104.2),
        ('R051', 'Benjamin Okonkwo',   'North', 'Inactive', 23100, 24000,  96.3),
        ('R052', 'Cassandra Wills',    'South', 'Inactive', 61400, 60000, 102.3),
        ('R053', 'Dominic Faber',      'East',  'Active',   34500, 33000, 104.5),
        ('R054', 'Eleanor Joung',      'West',  'Active',   28900, 28000, 103.2),
        ('R055', 'Felix Boateng',      'South', 'Active',   46300, 45000, 102.9),
        ('R056', 'Grace Lindqvist',    'South', 'Active',   39800, 38000, 104.7),
        ('R057', 'Harvey Stroud',      'East',  'Inactive', 52700, 50000, 105.4),
        ('R058', 'Ingrid Halvorsen',   'West',  'Active',   33200, 32000, 103.8),
        ('R059', 'Jackie Ramirez',     'North', 'On Leave', 29600, 29000, 102.1),
        ('R060', 'Kirk Nielsen',       'South', 'Active',   17800, 20000,  89.0),
        ('R061', 'Laura Kimura',       'East',  'Active',   60100, 58000, 103.6),
        ('R062', 'Mason Osei',         'West',  'Inactive', 44200, 43000, 102.8),
        ('R063', 'Natalia Volkov',     'North', 'Active',    7600, 10000,  76.0),
        ('R064', 'Oscar Iriarte',      'South', 'Active',   43700, 42000, 104.0),
        ('R065', 'Paige Lancaster',    'East',  'Active',   26800, 26000, 103.1),
        ('R066', 'Quentin Ashby',      'West',  'Active',   51900, 50000, 103.8),
        ('R067', 'Rosa Beaumont',      'North', 'Inactive', 31200, 30000, 104.0),
        ('R068', 'Steven Tanaka',      'South', 'On Leave', 58900, 57000, 103.3),
        ('R069', 'Tamara Espinosa',    'East',  'Active',   22300, 23000,  97.0),
        ('R070', 'Ulric Pedersen',     'West',  'Active',   40600, 39000, 104.1),
        ('R071', 'Vera Okafor',        'North', 'Active',    5200,  8000,  65.0),
        ('R072', 'William Fentress',   'South', 'Active',   36200, 35000, 103.4),
        ('R073', 'Xander Blum',        'East',  'Inactive', 48900, 47000, 104.0),
        ('R074', 'Yasmin Holbrook',    'West',  'Active',   27400, 27000, 101.5),
        ('R075', 'Zane Pemberton',     'North', 'Inactive', 55600, 54000, 103.0),
        ('R076', 'Amy Callahan',       'South', 'Active',   46800, 45000, 104.0),
        ('R077', 'Brian Sweeney',      'East',  'Active',   35700, 35000, 102.0),
        ('R078', 'Clara Hoffmann',     'West',  'Inactive', 62100, 60000, 103.5),
        ('R079', 'David Anand',        'North', 'Active',   18400, 20000,  92.0),
        ('R080', 'Emma Kowalczyk',     'South', 'Active',   53800, 52000, 103.5),
        ('R081', 'Frank Dumont',       'East',  'Active',   40200, 39000, 103.1),
        ('R082', 'Gina Watkins',       'West',  'Active',   29600, 29000, 102.1),
        ('R083', 'Howard Brennan',     'North', 'On Leave', 64700, 63000, 102.7),
        ('R084', 'Ivy Nakamura',       'South', 'Inactive', 33100, 33000, 100.3),
        ('R085', 'Jacob Ferreira',     'East',  'Active',   21700, 22000,  98.6),
        ('R086', 'Karen Mwangi',       'West',  'Active',   47200, 46000, 102.6),
        ('R087', 'Louis Garnier',      'North', 'Active',   12000, 15000,  80.0),
        ('R088', 'Monica Stravinsky',  'South', 'Active',   71200, 70000, 101.7),
        ('R089', 'Neil Drummond',      'East',  'On Leave', 30400, 30000, 101.3),
        ('R090', 'Olga Petrov',        'West',  'Active',   23900, 24000,  99.6),
    ]

    # Verify count
    matching = [r for r in data if r[3] == 'Active' and r[2] == 'North' and r[4] > 25000]
    print(f'Rows matching all 3 criteria (Active + North + Revenue>25000): {len(matching)}')

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Total data rows: {len(data)}')

create_initial()
