"""
Initial Setup: Add date validation to Review Date column
Task ID: calc_gg3_043
Domain: libreoffice_calc

Creates a project schedule spreadsheet with 50 tasks across 7 columns.
Column G (Review Date) has NO data validation - the task is to add it.
"""

import os
import shlex
import subprocess
import time
from datetime import date, timedelta
import random

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_043'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

random.seed(42)


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Schedule ---
    ws = wb.active
    ws.title = 'Schedule'

    headers = ['Task ID', 'Task Name', 'Owner', 'Priority', 'Start Date', 'End Date', 'Review Date']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_side = Side(style='thin', color='000000')
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14

    # Realistic project task data
    task_names = [
        'Requirements gathering', 'Stakeholder interviews', 'Market analysis report',
        'Technical feasibility study', 'Architecture design document', 'UI/UX wireframes',
        'Database schema design', 'API specification draft', 'Security audit plan',
        'Sprint planning session', 'Frontend prototype build', 'Backend service setup',
        'Authentication module', 'User registration flow', 'Dashboard analytics page',
        'Payment integration', 'Email notification system', 'Search functionality',
        'Data migration script', 'Load testing framework', 'CI/CD pipeline setup',
        'Code review guidelines', 'Unit test coverage report', 'Integration test suite',
        'Performance benchmark', 'Accessibility compliance check', 'Mobile responsive layout',
        'Localization framework', 'Error handling middleware', 'Logging infrastructure',
        'Monitoring dashboard setup', 'Incident response playbook', 'Disaster recovery plan',
        'User acceptance testing', 'Beta release preparation', 'Documentation update',
        'Training materials creation', 'Vendor evaluation report', 'License compliance review',
        'Budget reconciliation', 'Resource allocation plan', 'Risk assessment matrix',
        'Change management process', 'Quality assurance checklist', 'Deployment runbook',
        'Post-launch monitoring', 'Customer feedback analysis', 'Sprint retrospective',
        'Quarterly roadmap update', 'Annual review presentation',
    ]

    owners = [
        'Sarah Chen', 'Marcus Johnson', 'Elena Rodriguez', 'David Kim',
        'Priya Patel', 'James O\'Brien', 'Aisha Mohammed', 'Liam Fischer',
        'Mei-Ling Wu', 'Carlos Gutierrez', 'Rachel Thompson', 'Nikolai Petrov',
    ]

    priorities = ['High', 'Medium', 'Low', 'Critical']
    priority_weights = [30, 40, 20, 10]

    date_format = 'yyyy-mm-dd'

    for i in range(50):
        row = i + 2
        task_id_val = f'PRJ-{1001 + i}'
        task_name = task_names[i]
        owner = random.choice(owners)
        priority = random.choices(priorities, weights=priority_weights, k=1)[0]

        # Start dates spread across 2024
        start_offset = random.randint(0, 300)
        start_date = date(2024, 1, 1) + timedelta(days=start_offset)
        # End date 7-45 days after start
        duration = random.randint(7, 45)
        end_date = start_date + timedelta(days=duration)

        # Review dates: mostly in 2024, but some in 2023 or 2025 to show the problem
        if i % 8 == 0:
            # Occasional out-of-scope dates (2023)
            review_date = date(2023, random.randint(9, 12), random.randint(1, 28))
        elif i % 11 == 0:
            # Occasional out-of-scope dates (2025)
            review_date = date(2025, random.randint(1, 3), random.randint(1, 28))
        else:
            review_offset = random.randint(0, 350)
            review_date = date(2024, 1, 1) + timedelta(days=review_offset)

        ws.cell(row=row, column=1, value=task_id_val)
        ws.cell(row=row, column=2, value=task_name)
        ws.cell(row=row, column=3, value=owner)
        ws.cell(row=row, column=4, value=priority)

        start_cell = ws.cell(row=row, column=5, value=start_date)
        start_cell.number_format = date_format
        end_cell = ws.cell(row=row, column=6, value=end_date)
        end_cell.number_format = date_format
        review_cell = ws.cell(row=row, column=7, value=review_date)
        review_cell.number_format = date_format

    # Freeze header row
    ws.freeze_panes = 'A2'

    # --- Sheet 2: Summary ---
    ws2 = wb.create_sheet('Summary')
    ws2['A1'] = 'Project Schedule Summary'
    ws2['A1'].font = Font(size=14, bold=True)
    ws2['A3'] = 'Total Tasks:'
    ws2['B3'] = 50
    ws2['A4'] = 'Date Range:'
    ws2['B4'] = 'Jan 2024 - Dec 2024'
    ws2['A5'] = 'Last Updated:'
    ws2['B5'] = date(2024, 10, 15)
    ws2['B5'].number_format = date_format
    ws2['A7'] = 'Priority Breakdown'
    ws2['A7'].font = Font(bold=True)
    ws2['A8'] = 'Critical'
    ws2['A9'] = 'High'
    ws2['A10'] = 'Medium'
    ws2['A11'] = 'Low'
    ws2['B8'] = '=COUNTIF(Schedule!D2:D51,"Critical")'
    ws2['B9'] = '=COUNTIF(Schedule!D2:D51,"High")'
    ws2['B10'] = '=COUNTIF(Schedule!D2:D51,"Medium")'
    ws2['B11'] = '=COUNTIF(Schedule!D2:D51,"Low")'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
