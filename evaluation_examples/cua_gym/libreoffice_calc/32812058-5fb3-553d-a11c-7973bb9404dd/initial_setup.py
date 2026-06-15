"""
Initial Setup: Create a sales data spreadsheet for pivot table task.
Task ID: calc_pivot_001
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_001'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'SalesData'

    # --- Headers ---
    headers = ['OrderID', 'Date', 'Category', 'Product', 'Quantity', 'UnitPrice', 'Revenue']
    header_font = Font(bold=True, size=11, name='Calibri')
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    thin_side = Side(style="thin", color="000000")
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, name='Calibri', color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12

    # --- Products per category with realistic prices ---
    products = {
        'Electronics': [
            ('Wireless Mouse', 29.99), ('USB-C Hub', 45.00), ('Bluetooth Speaker', 65.00),
            ('Laptop Stand', 39.99), ('Webcam HD', 79.99), ('Mechanical Keyboard', 89.99),
            ('HDMI Cable', 12.99), ('Power Bank', 35.00), ('Monitor Light Bar', 55.00),
            ('Noise-Cancelling Earbuds', 110.00),
        ],
        'Clothing': [
            ('Cotton T-Shirt', 24.99), ('Denim Jeans', 59.99), ('Running Shoes', 85.00),
            ('Winter Jacket', 120.00), ('Wool Scarf', 29.99), ('Casual Polo', 34.99),
            ('Athletic Shorts', 28.00), ('Dress Shirt', 49.99), ('Leather Belt', 22.00),
            ('Baseball Cap', 18.99),
        ],
        'Food': [
            ('Organic Coffee Beans', 14.99), ('Dark Chocolate Bar', 5.99),
            ('Mixed Nuts Pack', 9.99), ('Green Tea Box', 7.50),
            ('Olive Oil 500ml', 12.50), ('Protein Bar Pack', 18.99),
            ('Dried Mango Slices', 6.50), ('Almond Butter', 11.99),
            ('Sparkling Water 12pk', 8.99), ('Granola Cereal', 5.50),
        ],
        'Books': [
            ('Python Programming', 42.99), ('Data Science Handbook', 38.50),
            ('Machine Learning Intro', 35.00), ('Clean Code', 32.99),
            ('Design Patterns', 44.99), ('The Pragmatic Programmer', 40.00),
            ('Algorithms 4th Ed', 55.00), ('Web Development Basics', 28.99),
            ('Cloud Architecture', 48.00), ('DevOps Handbook', 36.50),
        ],
    }

    # Revenue targets
    targets = {
        'Electronics': 45200.0,
        'Clothing': 32100.0,
        'Food': 18700.0,
        'Books': 12400.0,
    }
    # Grand total = 108400

    # Distribute 200 rows across categories proportionally to targets
    # Electronics ~42%, Clothing ~30%, Food ~17%, Books ~11%
    category_row_counts = {
        'Electronics': 70,
        'Clothing': 55,
        'Food': 45,
        'Books': 30,
    }
    # total = 200

    months_2024 = [
        '2024-01-', '2024-02-', '2024-03-', '2024-04-',
        '2024-05-', '2024-06-', '2024-07-', '2024-08-',
        '2024-09-', '2024-10-', '2024-11-', '2024-12-',
    ]

    all_rows = []

    for category, row_count in category_row_counts.items():
        target_revenue = targets[category]
        cat_products = products[category]

        # Generate row_count - 1 rows with random revenue, last row is adjustment
        rows_data = []
        running_revenue = 0.0

        for i in range(row_count - 1):
            product_name, base_price = random.choice(cat_products)
            # Vary quantity 1-5
            qty = random.randint(1, 5)
            # Vary price slightly
            unit_price = round(base_price * random.uniform(0.9, 1.1), 2)
            revenue = round(qty * unit_price, 2)
            running_revenue += revenue

            month = random.choice(months_2024)
            day = random.randint(1, 28)
            date_str = f"{month}{day:02d}"

            rows_data.append((date_str, category, product_name, qty, unit_price, revenue))

        # Adjustment row to hit exact target
        remaining = round(target_revenue - running_revenue, 2)
        product_name, base_price = cat_products[0]
        if remaining > 0:
            adj_qty = 1
            adj_price = remaining
        else:
            # Edge case: overshoot. Use negative-ish approach (shouldn't happen with our ranges)
            adj_qty = 1
            adj_price = remaining

        adj_month = random.choice(months_2024)
        adj_day = random.randint(1, 28)
        adj_date = f"{adj_month}{adj_day:02d}"
        rows_data.append((adj_date, category, product_name, adj_qty, adj_price, remaining))

        all_rows.extend(rows_data)

    # Shuffle all rows so categories are mixed
    random.shuffle(all_rows)

    # Write data rows
    for idx, (date_str, category, product_name, qty, unit_price, revenue) in enumerate(all_rows):
        row_num = idx + 2  # row 2 onwards
        order_id = idx + 1

        ws.cell(row=row_num, column=1, value=order_id)
        ws.cell(row=row_num, column=2, value=date_str)
        ws.cell(row=row_num, column=3, value=category)
        ws.cell(row=row_num, column=4, value=product_name)
        ws.cell(row=row_num, column=5, value=qty)
        ws.cell(row=row_num, column=6, value=unit_price)
        ws.cell(row=row_num, column=7, value=revenue)

        # Number formatting
        ws.cell(row=row_num, column=6).number_format = '#,##0.00'
        ws.cell(row=row_num, column=7).number_format = '#,##0.00'

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto-filter on all data
    ws.auto_filter.ref = f'A1:G{len(all_rows) + 1}'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Verify sums
    import openpyxl as opx
    wb_check = opx.load_workbook(OUTPUT)
    ws_check = wb_check['SalesData']
    sums = {}
    for r in range(2, ws_check.max_row + 1):
        cat = ws_check.cell(row=r, column=3).value
        rev = ws_check.cell(row=r, column=7).value
        if cat:
            sums[cat] = sums.get(cat, 0) + rev
    for cat, total in sorted(sums.items()):
        print(f'  {cat}: {total:.2f}')
    print(f'  Grand Total: {sum(sums.values()):.2f}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
