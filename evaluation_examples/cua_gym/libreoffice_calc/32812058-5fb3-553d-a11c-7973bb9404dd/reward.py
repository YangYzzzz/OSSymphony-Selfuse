"""
Reward Script: Create pivot table from sales data showing total revenue per product category
Task ID: calc_pivot_001
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): PivotTable sheet exists
  Component 2 (0.15): Correct header row (Category + Sum of Revenue)
  Component 3 (0.40): Revenue values per category match expected (4 x 0.10)
  Component 4 (0.20): Grand Total row with correct sum
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_001'

# Expected category revenue values from task context
EXPECTED_REVENUES = {
    'Electronics': 45200,
    'Clothing': 32100,
    'Food': 18700,
    'Books': 12400,
}
EXPECTED_GRAND_TOTAL = 108400


def find_pivot_sheet(wb):
    """Find the pivot table sheet (any sheet that is NOT 'SalesData')."""
    for name in wb.sheetnames:
        if name.lower() != 'salesdata':
            return wb[name]
    return None


def find_header_row(ws):
    """Find the row containing 'Category' header. Returns row number or None.

    We look for a row where one cell is exactly or closely 'Category' (not a title
    like 'Revenue by Category') and another cell contains 'revenue' or 'sum'.
    """
    for row in ws.iter_rows(min_row=1, max_row=min(15, ws.max_row), max_col=5):
        cells = {cell.column: (cell.value or '') for cell in row if cell.value and isinstance(cell.value, str)}
        has_category_header = any(
            v.strip().lower() == 'category' for v in cells.values()
        )
        has_revenue_header = any(
            'revenue' in v.lower() or 'sum' in v.lower() for v in cells.values()
        )
        if has_category_header and has_revenue_header:
            return row[0].row
    # Fallback: look for a cell that is exactly 'Category'
    for row in ws.iter_rows(min_row=1, max_row=min(15, ws.max_row), max_col=5):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.strip().lower() == 'category':
                return cell.row
    return None


def find_category_col_and_value_col(ws, header_row):
    """Find which columns contain category names and revenue values."""
    cat_col = None
    val_col = None
    for col in range(1, ws.max_column + 1):
        hdr = ws.cell(row=header_row, column=col).value
        if hdr and isinstance(hdr, str):
            hdr_lower = hdr.lower()
            if 'category' in hdr_lower:
                cat_col = col
            elif 'revenue' in hdr_lower or 'sum' in hdr_lower:
                val_col = col
    return cat_col, val_col


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: PivotTable sheet exists (0.25 points)
    # This FAILS on initial_env (only SalesData) and PASSES on golden_env
    try:
        pivot_ws = find_pivot_sheet(wb)
        if pivot_ws is not None:
            print(f"PASS: Component 1 -- Pivot sheet '{pivot_ws.title}' exists (0.25 pts)")
            total_score += 0.25
        else:
            print("FAIL: Component 1 -- No pivot table sheet found (only SalesData)")
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Component 2: Header row with Category and Revenue columns (0.15 points)
    # This FAILS on initial_env (no pivot sheet) and PASSES on golden_env
    try:
        header_row = find_header_row(pivot_ws)
        if header_row is not None:
            cat_col, val_col = find_category_col_and_value_col(pivot_ws, header_row)
            if cat_col is not None and val_col is not None:
                cat_hdr = pivot_ws.cell(row=header_row, column=cat_col).value
                val_hdr = pivot_ws.cell(row=header_row, column=val_col).value
                print(f"PASS: Component 2 -- Headers found at row {header_row}: "
                      f"'{cat_hdr}' (col {cat_col}), '{val_hdr}' (col {val_col}) (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 2 -- Header row found at {header_row} but missing "
                      f"Category col ({cat_col}) or Revenue col ({val_col})")
        else:
            print("FAIL: Component 2 -- No header row with 'Category' found in pivot sheet")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Revenue per category matches expected values (0.40 points, 0.10 each)
    # This FAILS on initial_env (no pivot sheet) and PASSES on golden_env
    try:
        if header_row is not None and cat_col is not None and val_col is not None:
            # Read all category-value pairs from pivot table
            found_categories = {}
            for r in range(header_row + 1, pivot_ws.max_row + 1):
                cat_val = pivot_ws.cell(row=r, column=cat_col).value
                rev_val = pivot_ws.cell(row=r, column=val_col).value
                if cat_val and isinstance(cat_val, str):
                    # Skip grand total row
                    if 'total' in cat_val.lower() or 'grand' in cat_val.lower():
                        continue
                    found_categories[cat_val.strip()] = rev_val

            for category, expected_rev in EXPECTED_REVENUES.items():
                if category in found_categories:
                    actual_rev = found_categories[category]
                    if actual_rev is not None:
                        try:
                            if abs(float(actual_rev) - expected_rev) <= 1.0:
                                print(f"PASS: Component 3 -- {category} revenue = {actual_rev} "
                                      f"(expected {expected_rev}) (0.10 pts)")
                                total_score += 0.10
                            else:
                                print(f"FAIL: Component 3 -- {category} revenue = {actual_rev}, "
                                      f"expected {expected_rev}")
                        except (ValueError, TypeError):
                            print(f"FAIL: Component 3 -- {category} revenue not numeric: {actual_rev}")
                    else:
                        print(f"FAIL: Component 3 -- {category} revenue is None")
                else:
                    print(f"FAIL: Component 3 -- Category '{category}' not found in pivot table. "
                          f"Found: {list(found_categories.keys())}")
        else:
            print("FAIL: Component 3 -- Cannot check revenues (no valid headers)")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Grand Total row with correct sum (0.20 points)
    # This FAILS on initial_env (no pivot sheet) and PASSES on golden_env
    try:
        if header_row is not None and cat_col is not None and val_col is not None:
            grand_total_found = False
            for r in range(header_row + 1, pivot_ws.max_row + 1):
                cat_val = pivot_ws.cell(row=r, column=cat_col).value
                if cat_val and isinstance(cat_val, str) and (
                    'total' in cat_val.lower() or 'grand' in cat_val.lower()
                ):
                    rev_val = pivot_ws.cell(row=r, column=val_col).value
                    if rev_val is not None:
                        try:
                            if abs(float(rev_val) - EXPECTED_GRAND_TOTAL) <= 1.0:
                                print(f"PASS: Component 4 -- Grand Total = {rev_val} "
                                      f"(expected {EXPECTED_GRAND_TOTAL}) (0.20 pts)")
                                total_score += 0.20
                                grand_total_found = True
                            else:
                                print(f"FAIL: Component 4 -- Grand Total = {rev_val}, "
                                      f"expected {EXPECTED_GRAND_TOTAL}")
                                grand_total_found = True
                        except (ValueError, TypeError):
                            print(f"FAIL: Component 4 -- Grand Total not numeric: {rev_val}")
                            grand_total_found = True
                    break
            if not grand_total_found:
                print("FAIL: Component 4 -- No Grand Total row found in pivot table")
        else:
            print("FAIL: Component 4 -- Cannot check grand total (no valid headers)")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
