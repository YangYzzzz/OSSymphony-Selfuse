"""
Reward Script: Build a discount approval matrix using nested IF formulas
Task ID: calc_sales_029
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): All 6 cells D2:D7 contain IF formulas
  Component 2 (0.5): Formulas produce correct approval levels for each deal
  Component 3 (0.2): Formulas are properly nested IFs referencing column C
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_029'

# Ground truth: discount % -> expected approval
# D2: 5% -> Auto-Approved, D3: 15% -> Manager, D4: 25% -> Director
# D5: 8% -> Auto-Approved, D6: 35% -> VP, D7: 22% -> Director
EXPECTED_APPROVALS = {
    2: ('Auto-Approved', 0.05),
    3: ('Manager', 0.15),
    4: ('Director', 0.25),
    5: ('Auto-Approved', 0.08),
    6: ('VP', 0.35),
    7: ('Director', 0.22),
}


def evaluate_nested_if(formula_str, discount_pct):
    """
    Evaluate the nested IF formula logic to determine the output string.
    Parses the formula to extract thresholds and labels, then applies logic.
    Returns the evaluated result string, or None if parsing fails.
    """
    # Normalize: strip leading =, uppercase
    f = formula_str.strip()
    if f.startswith('='):
        f = f[1:]

    # Extract all IF conditions and string literals from the formula
    # Pattern: IF(Cx<=threshold,"Label", ...) or IF(Cx<threshold,...) etc.
    # We'll simulate by extracting threshold-label pairs
    # Find all patterns like: C\d+<=0.1,"Auto-Approved"  or C\d+<0.1,"Auto-Approved"
    pairs = re.findall(r'C\d+\s*([<>]=?)\s*([\d.]+)\s*,\s*"([^"]+)"', f, re.IGNORECASE)

    if not pairs:
        return None

    # Build rules: list of (operator, threshold, label)
    rules = []
    for op, thresh, label in pairs:
        rules.append((op, float(thresh), label))

    # The last string literal that isn't part of a condition is the else case
    all_strings = re.findall(r'"([^"]+)"', f)
    # Labels already captured in rules
    rule_labels = [label for _, _, label in rules]
    # The else value is the last string that isn't a rule label at its position
    # Typically the very last quoted string
    else_label = all_strings[-1] if all_strings else None

    # If the else_label is the same as the last rule's label, then the actual
    # else is the remaining one
    # In a nested IF like IF(C<=0.1,"A",IF(C<=0.2,"B",IF(C<=0.3,"C","D")))
    # all_strings = ["A","B","C","D"], rules capture A,B,C, else is D
    # But if there's an extra nesting, else_label might be a rule label
    # Simple approach: evaluate sequentially
    for op, thresh, label in rules:
        if op == '<=' and discount_pct <= thresh:
            return label
        elif op == '<' and discount_pct < thresh:
            return label
        elif op == '>=' and discount_pct >= thresh:
            return label
        elif op == '>' and discount_pct > thresh:
            return label

    return else_label


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists
    if 'DiscountApproval' not in wb.sheetnames:
        print("FAIL: Sheet 'DiscountApproval' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['DiscountApproval']

    # Component 1: All 6 cells D2:D7 contain IF formulas (0.3 points)
    try:
        formula_count = 0
        for row in range(2, 8):
            val = ws.cell(row=row, column=4).value
            if val is not None and isinstance(val, str) and 'IF(' in val.upper():
                formula_count += 1
            else:
                print(f"  D{row}: not an IF formula, found: {val}")

        if formula_count == 6:
            print(f"PASS: Component 1 -- All 6 cells D2:D7 contain IF formulas (0.3 pts)")
            total_score += 0.3
        elif formula_count > 0:
            partial = round(0.3 * (formula_count / 6), 2)
            print(f"PARTIAL: Component 1 -- {formula_count}/6 cells have IF formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 -- No IF formulas found in D2:D7")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Formulas evaluate to correct ground truth values (0.5 points)
    # Each correct cell is worth 0.5/6 ~ 0.083 points
    try:
        correct_count = 0
        per_cell = round(0.5 / 6, 4)
        for row, (expected_label, discount_pct) in EXPECTED_APPROVALS.items():
            val = ws.cell(row=row, column=4).value
            if val is None:
                print(f"  D{row}: empty, expected formula evaluating to '{expected_label}'")
                continue

            if isinstance(val, str) and val.startswith('='):
                # It's a formula - evaluate it
                result = evaluate_nested_if(val, discount_pct)
                if result == expected_label:
                    correct_count += 1
                    print(f"  D{row}: formula evaluates to '{result}' == '{expected_label}' OK")
                else:
                    print(f"  D{row}: formula evaluates to '{result}', expected '{expected_label}'")
            elif isinstance(val, str) and val == expected_label:
                # It's a plain text value matching expected (e.g., cached value)
                correct_count += 1
                print(f"  D{row}: plain value '{val}' == '{expected_label}' OK")
            else:
                print(f"  D{row}: value '{val}' does not match expected '{expected_label}'")

        if correct_count == 6:
            print(f"PASS: Component 2 -- All 6 formulas evaluate correctly (0.5 pts)")
            total_score += 0.5
        elif correct_count > 0:
            comp2_score = round(per_cell * correct_count, 2)
            print(f"PARTIAL: Component 2 -- {correct_count}/6 formulas correct ({comp2_score} pts)")
            total_score += comp2_score
        else:
            print(f"FAIL: Component 2 -- No formulas evaluate correctly")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Formulas are properly nested IFs referencing column C (0.2 points)
    # Check that formulas use nested IF structure and reference the C column cells
    try:
        nested_count = 0
        for row in range(2, 8):
            val = ws.cell(row=row, column=4).value
            if val is None or not isinstance(val, str):
                continue
            upper_val = val.upper().replace(' ', '')
            # Count IF occurrences - nested IF should have at least 3 IFs
            if_count = upper_val.count('IF(')
            # Check reference to column C for this row
            has_c_ref = f'C{row}' in val.upper().replace(' ', '')
            if if_count >= 3 and has_c_ref:
                nested_count += 1
            else:
                print(f"  D{row}: IF count={if_count}, C{row} ref={has_c_ref}")

        if nested_count == 6:
            print(f"PASS: Component 3 -- All formulas use nested IFs with C column refs (0.2 pts)")
            total_score += 0.2
        elif nested_count > 0:
            partial = round(0.2 * (nested_count / 6), 2)
            print(f"PARTIAL: Component 3 -- {nested_count}/6 formulas properly nested ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 -- No properly nested IF formulas found")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
