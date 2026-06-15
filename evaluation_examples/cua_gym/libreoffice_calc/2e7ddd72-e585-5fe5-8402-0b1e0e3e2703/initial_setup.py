"""
Initial Setup: Discount approval matrix spreadsheet with deal data
Task ID: calc_sales_029
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_029'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: DiscountApproval ---
    ws = wb.active
    ws.title = 'DiscountApproval'

    # Headers
    headers = ['Deal', 'Value', 'Requested Discount %', 'Approval Required']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows - deals with discount percentages
    # Column D (Approval Required) is intentionally LEFT EMPTY for the agent to fill
    data = [
        ['Deal A', 50000, 0.05],
        ['Deal B', 120000, 0.15],
        ['Deal C', 200000, 0.25],
        ['Deal D', 75000, 0.08],
        ['Deal E', 300000, 0.35],
        ['Deal F', 90000, 0.22],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 1:
                cell.font = Font(name='Calibri', size=11)
            elif c == 2:
                cell.number_format = '$#,##0'
                cell.alignment = Alignment(horizontal='right')
            elif c == 3:
                cell.number_format = '0%'
                cell.alignment = Alignment(horizontal='center')
        # D column empty but bordered
        d_cell = ws.cell(row=r, column=4)
        d_cell.border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 22

    # Freeze header row
    ws.freeze_panes = 'A2'

    # --- Sheet: ApprovalPolicy ---
    ws2 = wb.create_sheet('ApprovalPolicy')
    ws2['A1'] = 'Discount Threshold'
    ws2['B1'] = 'Approver Level'
    ws2['A1'].font = Font(bold=True)
    ws2['B1'].font = Font(bold=True)

    policy_data = [
        ['Up to 10%', 'Auto-Approved'],
        ['10% - 20%', 'Manager'],
        ['20% - 30%', 'Director'],
        ['Over 30%', 'VP'],
    ]
    for r, row_data in enumerate(policy_data, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
