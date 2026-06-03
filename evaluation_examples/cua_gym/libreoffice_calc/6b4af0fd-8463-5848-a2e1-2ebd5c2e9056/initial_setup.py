"""
Initial Setup: Create real estate property listing data for pivot table task
Task ID: calc_pivot_080
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_080'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Properties'

    # Headers
    headers = ['ListingID', 'Neighborhood', 'PropertyType', 'SqFt', 'Price', 'PricePerSqFt']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    neighborhoods = ['Downtown', 'Suburbs', 'Midtown', 'Waterfront', 'University']
    property_types = ['Apartment', 'House', 'Condo', 'Townhouse']

    # We need to engineer data so that:
    #   avg PricePerSqFt for Downtown/Apartment = 450
    #   avg PricePerSqFt for Suburbs/House = 185
    #   grand total avg PricePerSqFt across all 180 rows = 280
    #
    # Strategy: assign PricePerSqFt ranges by neighborhood+type combo,
    # then compute Price = PricePerSqFt * SqFt

    # Define target avg PricePerSqFt for each (neighborhood, property_type) combo
    # We'll create 9 rows per combo (180 / 20 combos = 9 rows each)
    target_avg = {
        ('Downtown', 'Apartment'): 450,
        ('Downtown', 'House'): 380,
        ('Downtown', 'Condo'): 410,
        ('Downtown', 'Townhouse'): 390,
        ('Suburbs', 'Apartment'): 160,
        ('Suburbs', 'House'): 185,
        ('Suburbs', 'Condo'): 170,
        ('Suburbs', 'Townhouse'): 175,
        ('Midtown', 'Apartment'): 320,
        ('Midtown', 'House'): 270,
        ('Midtown', 'Condo'): 280,
        ('Midtown', 'Townhouse'): 280,
        ('Waterfront', 'Apartment'): 380,
        ('Waterfront', 'House'): 310,
        ('Waterfront', 'Condo'): 330,
        ('Waterfront', 'Townhouse'): 320,
        ('University', 'Apartment'): 210,
        ('University', 'House'): 195,
        ('University', 'Condo'): 195,
        ('University', 'Townhouse'): 190,
    }

    # Verify grand total average = 280
    all_avgs = list(target_avg.values())
    grand_avg = sum(all_avgs) / len(all_avgs)
    print(f"Design grand average: {grand_avg}")  # Should be 280

    rows_per_combo = 9  # 20 combos * 9 = 180 rows
    data = []

    for (neighborhood, ptype), avg_ppsf in target_avg.items():
        # Generate rows_per_combo values that average to avg_ppsf
        # Use symmetric spread around the mean
        values = []
        for i in range(rows_per_combo):
            # Spread +/- 15% around mean, ensuring mean is exact
            offset = (i - (rows_per_combo - 1) / 2) * (avg_ppsf * 0.03)
            val = round(avg_ppsf + offset, 2)
            values.append(val)

        # Adjust to ensure exact average
        current_avg = sum(values) / len(values)
        diff = avg_ppsf - current_avg
        values = [round(v + diff, 2) for v in values]

        for ppsf in values:
            sqft = random.randint(600, 3500)
            price = round(ppsf * sqft, 2)
            data.append([neighborhood, ptype, sqft, price, ppsf])

    # Shuffle data for realism
    random.shuffle(data)

    # Write data rows
    for idx, (neighborhood, ptype, sqft, price, ppsf) in enumerate(data, 1):
        row = idx + 1  # row 2 onwards
        ws.cell(row=row, column=1, value=idx)           # ListingID
        ws.cell(row=row, column=2, value=neighborhood)   # Neighborhood
        ws.cell(row=row, column=3, value=ptype)          # PropertyType
        ws.cell(row=row, column=4, value=sqft)           # SqFt
        ws.cell(row=row, column=5, value=price)          # Price
        ws.cell(row=row, column=6, value=ppsf)           # PricePerSqFt

    # Format Price column as currency
    for r in range(2, 182):
        ws.cell(row=r, column=5).number_format = '$#,##0.00'

    # Format PricePerSqFt as number with 2 decimals
    for r in range(2, 182):
        ws.cell(row=r, column=6).number_format = '#,##0.00'

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14

    # Bold headers
    from openpyxl.styles import Font
    for col in range(1, 7):
        ws.cell(row=1, column=col).font = Font(bold=True)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Verify data integrity
    print(f'Total data rows: {len(data)}')

    # Quick check expected averages
    from collections import defaultdict
    combo_sums = defaultdict(list)
    for neighborhood, ptype, sqft, price, ppsf in data:
        combo_sums[(neighborhood, ptype)].append(ppsf)

    dt_apt = combo_sums[('Downtown', 'Apartment')]
    print(f"Downtown/Apartment avg: {sum(dt_apt)/len(dt_apt):.2f}")
    sub_house = combo_sums[('Suburbs', 'House')]
    print(f"Suburbs/House avg: {sum(sub_house)/len(sub_house):.2f}")
    all_vals = [v for vals in combo_sums.values() for v in vals]
    print(f"Grand total avg: {sum(all_vals)/len(all_vals):.2f}")

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
