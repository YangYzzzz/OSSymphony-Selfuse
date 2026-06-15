"""
Reward Script: Create a pivot table from real estate data with average price per sq ft
                by neighborhood (rows) and property type (columns).
Task ID: calc_pivot_080
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Pivot table structure exists (data beyond original 6 columns)
  Component 2 (0.20): Correct neighborhood row labels (Downtown, Midtown, Suburbs, University, Waterfront)
  Component 3 (0.20): Correct property type column labels (Apartment, Condo, House, Townhouse)
  Component 4 (0.20): Key values correct (Downtown/Apartment~450, Suburbs/House~185)
  Component 5 (0.15): Grand total average ~280
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_080'

EXPECTED_NEIGHBORHOODS = {'Downtown', 'Midtown', 'Suburbs', 'University', 'Waterfront'}
EXPECTED_PROPERTY_TYPES = {'Apartment', 'Condo', 'House', 'Townhouse'}


def find_pivot_area(wb):
    """
    Search all sheets for the pivot table. It could be on a new sheet or
    appended to the existing Properties sheet (beyond column F).
    Returns (ws, header_row, data_start_row, row_label_col, first_data_col, last_data_col)
    or None if not found.
    """
    for ws in wb.worksheets:
        max_col = ws.max_column
        max_row = ws.max_row
        # Scan all cells looking for "Neighborhood" as a label in the pivot header
        for r in range(1, min(max_row + 1, 30)):
            for c in range(1, max_col + 1):
                val = ws.cell(row=r, column=c).value
                if val and str(val).strip().lower() == 'neighborhood':
                    # Found the row label header. Check if property types are in the same row
                    row_types = {}
                    for cc in range(c + 1, min(max_col + 1, c + 20)):
                        cv = ws.cell(row=r, column=cc).value
                        if cv and str(cv).strip() in EXPECTED_PROPERTY_TYPES:
                            row_types[str(cv).strip()] = cc
                    if len(row_types) >= 2:
                        # Found pivot area
                        # Determine last data column (look for Total Result or last property type col)
                        last_data_col = max(row_types.values())
                        # Check one more column for Total Result
                        for cc in range(last_data_col + 1, last_data_col + 3):
                            cv = ws.cell(row=r, column=cc).value
                            if cv and 'total' in str(cv).strip().lower():
                                last_data_col = cc
                                break
                        return {
                            'ws': ws,
                            'header_row': r,
                            'row_label_col': c,
                            'property_type_cols': row_types,
                            'last_data_col': last_data_col,
                        }
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Pivot table structure exists (0.25 points)
    # The pivot table should have data beyond the original 6 columns (A-F),
    # either on the same sheet or a new sheet, with neighborhood rows and
    # property type columns.
    try:
        pivot = find_pivot_area(wb)
        if pivot is not None:
            print(f"PASS: Component 1 — Pivot table structure found on sheet '{pivot['ws'].title}', "
                  f"header row {pivot['header_row']}, {len(pivot['property_type_cols'])} property type columns (0.25 pts)")
            total_score += 0.25
        else:
            print("FAIL: Component 1 — No pivot table structure found (no 'Neighborhood' row label "
                  "with property type column headers)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    if pivot is None:
        # Cannot proceed without pivot area
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    ws = pivot['ws']
    header_row = pivot['header_row']
    row_label_col = pivot['row_label_col']
    pt_cols = pivot['property_type_cols']
    last_data_col = pivot['last_data_col']

    # Component 2: Correct neighborhood row labels (0.20 points)
    # Scan rows below the header for the 5 expected neighborhoods
    try:
        found_neighborhoods = {}
        for r in range(header_row + 1, header_row + 20):
            val = ws.cell(row=r, column=row_label_col).value
            if val and str(val).strip() in EXPECTED_NEIGHBORHOODS:
                found_neighborhoods[str(val).strip()] = r
        match_count = len(found_neighborhoods)
        if match_count == 5:
            print(f"PASS: Component 2 — All 5 neighborhoods found as row labels: {sorted(found_neighborhoods.keys())} (0.20 pts)")
            total_score += 0.20
        elif match_count >= 3:
            partial = round(0.20 * match_count / 5, 2)
            print(f"PARTIAL: Component 2 — {match_count}/5 neighborhoods found: {sorted(found_neighborhoods.keys())} ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {match_count}/5 neighborhoods found: {sorted(found_neighborhoods.keys())}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Correct property type column labels (0.20 points)
    try:
        found_types = set(pt_cols.keys())
        missing = EXPECTED_PROPERTY_TYPES - found_types
        if len(found_types) == 4:
            print(f"PASS: Component 3 — All 4 property types found as column labels: {sorted(found_types)} (0.20 pts)")
            total_score += 0.20
        elif len(found_types) >= 2:
            partial = round(0.20 * len(found_types) / 4, 2)
            print(f"PARTIAL: Component 3 — {len(found_types)}/4 property types found: {sorted(found_types)}, missing: {sorted(missing)} ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {len(found_types)}/4 property types found: {sorted(found_types)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Key values correct (0.20 points)
    # Downtown/Apartment should be ~450, Suburbs/House should be ~185
    # Each sub-check is worth 0.10
    try:
        checks_passed = 0

        # Check Downtown/Apartment ~450
        if 'Apartment' in pt_cols and 'Downtown' in found_neighborhoods:
            dt_apt_val = ws.cell(row=found_neighborhoods['Downtown'], column=pt_cols['Apartment']).value
            if dt_apt_val is not None:
                try:
                    dt_apt_num = float(dt_apt_val)
                    if abs(dt_apt_num - 450) <= 15:
                        print(f"PASS: Component 4a — Downtown/Apartment = {dt_apt_num} (expected ~450)")
                        checks_passed += 1
                    else:
                        print(f"FAIL: Component 4a — Downtown/Apartment = {dt_apt_num} (expected ~450)")
                except (ValueError, TypeError):
                    print(f"FAIL: Component 4a — Downtown/Apartment value is not numeric: {dt_apt_val}")
            else:
                print("FAIL: Component 4a — Downtown/Apartment cell is empty")
        else:
            print("FAIL: Component 4a — Cannot locate Downtown row or Apartment column")

        # Check Suburbs/House ~185
        if 'House' in pt_cols and 'Suburbs' in found_neighborhoods:
            sub_house_val = ws.cell(row=found_neighborhoods['Suburbs'], column=pt_cols['House']).value
            if sub_house_val is not None:
                try:
                    sub_house_num = float(sub_house_val)
                    if abs(sub_house_num - 185) <= 15:
                        print(f"PASS: Component 4b — Suburbs/House = {sub_house_num} (expected ~185)")
                        checks_passed += 1
                    else:
                        print(f"FAIL: Component 4b — Suburbs/House = {sub_house_num} (expected ~185)")
                except (ValueError, TypeError):
                    print(f"FAIL: Component 4b — Suburbs/House value is not numeric: {sub_house_val}")
            else:
                print("FAIL: Component 4b — Suburbs/House cell is empty")
        else:
            print("FAIL: Component 4b — Cannot locate Suburbs row or House column")

        sub_score = round(0.20 * checks_passed / 2, 2)
        if checks_passed > 0:
            print(f"Component 4 subtotal: {checks_passed}/2 checks passed ({sub_score} pts)")
        total_score += sub_score
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Grand total average ~280 (0.15 points)
    # Look for a "Total Result" or "Grand Total" row
    try:
        grand_total_found = False
        for r in range(header_row + 1, header_row + 20):
            val = ws.cell(row=r, column=row_label_col).value
            if val and 'total' in str(val).strip().lower():
                # Found a total row - check the last column (overall average) or
                # look for a value near 280
                # Try the last data column first
                total_val = ws.cell(row=r, column=last_data_col).value
                if total_val is not None:
                    try:
                        total_num = float(total_val)
                        if abs(total_num - 280) <= 15:
                            print(f"PASS: Component 5 — Grand total average = {total_num} (expected ~280) (0.15 pts)")
                            total_score += 0.15
                            grand_total_found = True
                            break
                        else:
                            print(f"FAIL: Component 5 — Grand total average = {total_num} (expected ~280)")
                            grand_total_found = True
                            break
                    except (ValueError, TypeError):
                        pass
                # Also check cells in the property type columns for a value near 280
                for cc in pt_cols.values():
                    cv = ws.cell(row=r, column=cc).value
                    if cv is not None:
                        try:
                            cv_num = float(cv)
                            # These are per-type totals, not grand total
                        except (ValueError, TypeError):
                            pass

        if not grand_total_found:
            # Try to find grand total by looking for any cell with value ~280 in the total row area
            print("FAIL: Component 5 — No grand total row found with average ~280")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
import os

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
