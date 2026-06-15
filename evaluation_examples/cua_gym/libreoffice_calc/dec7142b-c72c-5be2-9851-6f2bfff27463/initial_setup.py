"""
Initial Setup: Set print area to A1:D20 while data extends to F35
Task ID: calc_tbl_046
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_046'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    # --- Headers (row 1) ---
    headers = ["Item Code", "Product Name", "Category", "Unit Price", "Qty in Stock", "Total Value"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Data rows 2-35 (34 rows of realistic inventory data) ---
    data = [
        ["INV-001", "Wireless Mouse", "Electronics", 29.99, 145, None],
        ["INV-002", "USB-C Hub 7-Port", "Electronics", 49.95, 82, None],
        ["INV-003", "Ergonomic Keyboard", "Electronics", 89.50, 63, None],
        ["INV-004", "Monitor Stand Riser", "Furniture", 34.99, 110, None],
        ["INV-005", "LED Desk Lamp", "Lighting", 42.00, 97, None],
        ["INV-006", "Noise-Cancel Headphones", "Electronics", 199.99, 38, None],
        ["INV-007", "Webcam HD 1080p", "Electronics", 74.50, 55, None],
        ["INV-008", "Standing Desk Mat", "Furniture", 45.00, 72, None],
        ["INV-009", "Cable Management Kit", "Accessories", 15.99, 230, None],
        ["INV-010", "Laptop Sleeve 15in", "Accessories", 24.99, 168, None],
        ["INV-011", "Mechanical Keyboard", "Electronics", 129.00, 41, None],
        ["INV-012", "Desk Organizer Tray", "Furniture", 22.50, 195, None],
        ["INV-013", "HDMI Cable 6ft", "Accessories", 12.99, 310, None],
        ["INV-014", "Portable SSD 1TB", "Electronics", 109.99, 27, None],
        ["INV-015", "Office Chair Cushion", "Furniture", 39.99, 86, None],
        ["INV-016", "Whiteboard 36x24", "Office Supplies", 54.00, 33, None],
        ["INV-017", "Dry Erase Markers 12pk", "Office Supplies", 8.99, 420, None],
        ["INV-018", "Surge Protector 8-Outlet", "Electronics", 27.50, 148, None],
        ["INV-019", "Desk Phone Stand", "Accessories", 18.99, 203, None],
        ["INV-020", "Bluetooth Speaker", "Electronics", 65.00, 59, None],
        ["INV-021", "Filing Cabinet 3-Drawer", "Furniture", 189.00, 14, None],
        ["INV-022", "Paper Shredder", "Office Supplies", 79.99, 22, None],
        ["INV-023", "Ethernet Cable 25ft", "Accessories", 9.99, 275, None],
        ["INV-024", "Webcam Ring Light", "Lighting", 32.00, 91, None],
        ["INV-025", "USB Flash Drive 64GB", "Electronics", 11.50, 350, None],
        ["INV-026", "Document Scanner", "Electronics", 249.99, 12, None],
        ["INV-027", "Desk Bookshelf", "Furniture", 67.00, 45, None],
        ["INV-028", "Anti-Glare Screen Filter", "Accessories", 28.99, 76, None],
        ["INV-029", "Label Maker", "Office Supplies", 39.50, 58, None],
        ["INV-030", "Power Strip Tower", "Electronics", 35.99, 132, None],
        ["INV-031", "Footrest Adjustable", "Furniture", 44.00, 64, None],
        ["INV-032", "Sticky Notes Variety Pack", "Office Supplies", 6.99, 510, None],
        ["INV-033", "Wireless Charger Pad", "Electronics", 19.99, 187, None],
        ["INV-034", "Privacy Screen 27in", "Accessories", 52.00, 29, None],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)

    # Column F: Total Value = Unit Price * Qty in Stock (plain values, not formulas)
    for r in range(2, 36):
        price = ws.cell(row=r, column=4).value
        qty = ws.cell(row=r, column=5).value
        if price is not None and qty is not None:
            ws.cell(row=r, column=6, value=round(price * qty, 2))

    # Number formatting
    for r in range(2, 36):
        ws.cell(row=r, column=4).number_format = '$#,##0.00'
        ws.cell(row=r, column=6).number_format = '$#,##0.00'

    # Column widths for readability
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14

    # --- Set print area to A1:D20 (intentionally incomplete) ---
    ws.print_area = "A1:D20"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
