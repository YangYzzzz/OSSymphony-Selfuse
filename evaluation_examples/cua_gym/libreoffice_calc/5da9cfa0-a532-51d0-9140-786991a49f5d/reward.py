"""
Reward Script: EEO-1 Headcount Summary with COUNTIFS formulas and borders
Task ID: calc_hr_eeo_report_011
Domain: libreoffice_calc
Scoring:
  Component 1: Title header A1 — merged, bold 14pt, centered, correct text (0.25 pts)
  Component 2: Column headers A3:D3 — correct labels, bold (0.15 pts)
  Component 3: Ethnicity labels A4:A9 — all 6 categories (0.15 pts)
  Component 4: COUNTIFS formulas B4:C9 — correct formula structure (0.25 pts)
  Component 5: Total row A10:D10 — label and SUM formulas (0.10 pts)
  Component 6: Borders A3:D10 — thin borders on all cells (0.10 pts)
  Total: 1.0
"""

import os
from openpyxl.cell.cell import MergedCell

try:
    import openpyxl
except ImportError:
    print("CRITICAL: openpyxl not installed")
    print("REWARD: 0.0")
    exit(0)

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_eeo_report_011'

EXPECTED_ETHNICITIES = ['White', 'Black', 'Hispanic', 'Asian', 'Two or More Races', 'Other']


def verify_task(file_path):
    """
    Verify EEO-1 headcount summary task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: EEO Summary sheet must exist
    if 'EEO Summary' not in wb.sheetnames:
        print("FAIL: Sheet 'EEO Summary' does not exist")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['EEO Summary']

    # Component 1: Title header (0.25 points)
    # A1 must be merged across A1:D1, contain 'EEO-1 Headcount Summary', bold 14pt, centered
    try:
        # Check merge
        merged_ranges = [str(mr) for mr in ws.merged_cells.ranges]
        is_merged = 'A1:D1' in merged_ranges

        # Check text
        a1_val = ws['A1'].value
        is_correct_text = a1_val and 'EEO-1 Headcount Summary' in str(a1_val)

        # Check bold and size
        is_bold = ws['A1'].font.bold == True
        font_size = ws['A1'].font.size
        is_correct_size = font_size is not None and float(font_size) >= 13.0  # 14pt expected

        # Check alignment
        align_h = ws['A1'].alignment.horizontal
        is_centered = align_h in ('center', 'centerContinuous')

        if is_merged and is_correct_text and is_bold and is_correct_size and is_centered:
            print(f"PASS: Component 1 — A1 title header correct: merged={is_merged}, "
                  f"text={repr(a1_val)}, bold={is_bold}, size={font_size}, align={align_h} (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — A1 title header incomplete: "
                  f"merged={is_merged}, text={repr(a1_val)}, bold={is_bold}, "
                  f"size={font_size}, centered={is_centered}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Column headers A3:D3 (0.15 points)
    # A3='Ethnicity', B3='Male', C3='Female', D3='Total', all bold
    try:
        expected_headers = {'A3': 'Ethnicity', 'B3': 'Male', 'C3': 'Female', 'D3': 'Total'}
        headers_correct = True
        headers_bold = True
        for coord, expected_val in expected_headers.items():
            cell = ws[coord]
            actual_val = cell.value
            if actual_val is None or str(actual_val).strip() != expected_val:
                headers_correct = False
                print(f"FAIL: Component 2 — {coord} expected '{expected_val}', got {repr(actual_val)}")
            if not cell.font.bold:
                headers_bold = False
                print(f"FAIL: Component 2 — {coord} expected bold=True, got bold={cell.font.bold}")

        if headers_correct and headers_bold:
            print(f"PASS: Component 2 — column headers A3:D3 correct and bold (0.15 pts)")
            total_score += 0.15
        elif headers_correct:
            print(f"PARTIAL: Component 2 — headers have correct text but not all bold")
        # No partial credit for this component — award full or nothing
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Ethnicity labels A4:A9 (0.15 points)
    # A4:A9 must contain exactly the 6 expected ethnicity categories in order
    try:
        actual_ethnicities = []
        for row_idx in range(4, 10):
            cell_val = ws.cell(row=row_idx, column=1).value
            actual_ethnicities.append(str(cell_val).strip() if cell_val else '')

        correct_count = sum(1 for actual, expected in zip(actual_ethnicities, EXPECTED_ETHNICITIES)
                           if actual == expected)

        if correct_count == 6:
            print(f"PASS: Component 3 — all 6 ethnicity labels correct in A4:A9 (0.15 pts)")
            total_score += 0.15
        elif correct_count >= 3:
            partial = 0.07
            print(f"PARTIAL: Component 3 — {correct_count}/6 ethnicity labels correct: "
                  f"{actual_ethnicities} (partial {partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — only {correct_count}/6 ethnicity labels correct: "
                  f"{actual_ethnicities}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: COUNTIFS formulas in B4:C9 (0.25 points)
    # B4:B9 = COUNTIFS for Male per ethnicity, C4:C9 = COUNTIFS for Female per ethnicity
    # D4:D9 = =B4+C4 style formulas
    try:
        countifs_score = 0.0

        # Check B4:B9 for Male COUNTIFS formulas
        male_formulas_ok = 0
        for row_idx in range(4, 10):
            cell_val = ws.cell(row=row_idx, column=2).value
            if cell_val and isinstance(cell_val, str):
                val_upper = cell_val.upper()
                if ('COUNTIFS' in val_upper and
                    'HR DATA' in val_upper.replace("'", '') and
                    '"MALE"' in val_upper and
                    '"ACTIVE"' in val_upper):
                    male_formulas_ok += 1
                else:
                    print(f"FAIL: Component 4 — B{row_idx} does not match expected COUNTIFS Male pattern: {repr(cell_val)}")
            else:
                print(f"FAIL: Component 4 — B{row_idx} expected COUNTIFS formula, got: {repr(cell_val)}")

        # Check C4:C9 for Female COUNTIFS formulas
        female_formulas_ok = 0
        for row_idx in range(4, 10):
            cell_val = ws.cell(row=row_idx, column=3).value
            if cell_val and isinstance(cell_val, str):
                val_upper = cell_val.upper()
                if ('COUNTIFS' in val_upper and
                    'HR DATA' in val_upper.replace("'", '') and
                    '"FEMALE"' in val_upper and
                    '"ACTIVE"' in val_upper):
                    female_formulas_ok += 1
                else:
                    print(f"FAIL: Component 4 — C{row_idx} does not match expected COUNTIFS Female pattern: {repr(cell_val)}")
            else:
                print(f"FAIL: Component 4 — C{row_idx} expected COUNTIFS formula, got: {repr(cell_val)}")

        # Check D4:D9 for addition formulas (=B4+C4 etc.)
        total_formulas_ok = 0
        for row_idx in range(4, 10):
            cell_val = ws.cell(row=row_idx, column=4).value
            if cell_val and isinstance(cell_val, str):
                # Should be like =B4+C4
                val_clean = cell_val.strip().upper().replace(' ', '')
                b_col = f'B{row_idx}'
                c_col = f'C{row_idx}'
                if b_col in val_clean and c_col in val_clean and '+' in val_clean:
                    total_formulas_ok += 1
                else:
                    print(f"FAIL: Component 4 — D{row_idx} expected addition formula, got: {repr(cell_val)}")
            else:
                print(f"FAIL: Component 4 — D{row_idx} expected formula, got: {repr(cell_val)}")

        total_formula_checks = male_formulas_ok + female_formulas_ok + total_formulas_ok
        max_checks = 18  # 6 Male + 6 Female + 6 Total

        if total_formula_checks == max_checks:
            countifs_score = 0.25
            print(f"PASS: Component 4 — all COUNTIFS/addition formulas correct: "
                  f"{male_formulas_ok}/6 Male, {female_formulas_ok}/6 Female, "
                  f"{total_formulas_ok}/6 D-col (0.25 pts)")
        elif total_formula_checks >= 12:
            countifs_score = 0.15
            print(f"PARTIAL: Component 4 — {total_formula_checks}/{max_checks} formulas correct "
                  f"({male_formulas_ok}/6 Male, {female_formulas_ok}/6 Female, "
                  f"{total_formulas_ok}/6 D-col) (0.15 pts)")
        elif total_formula_checks >= 6:
            countifs_score = 0.08
            print(f"PARTIAL: Component 4 — {total_formula_checks}/{max_checks} formulas correct (0.08 pts)")
        else:
            print(f"FAIL: Component 4 — only {total_formula_checks}/{max_checks} formulas correct")

        total_score += countifs_score
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Total row A10:D10 (0.10 points)
    # A10='Total', B10=SUM(B4:B9), C10=SUM(C4:C9), D10=SUM(D4:D9)
    try:
        a10_val = ws['A10'].value
        b10_val = ws['B10'].value
        c10_val = ws['C10'].value
        d10_val = ws['D10'].value

        a10_ok = bool(a10_val and str(a10_val).strip() == 'Total')
        b10_ok = bool(b10_val and isinstance(b10_val, str) and
                  'SUM' in b10_val.upper() and 'B4' in b10_val.upper() and 'B9' in b10_val.upper())
        c10_ok = bool(c10_val and isinstance(c10_val, str) and
                  'SUM' in c10_val.upper() and 'C4' in c10_val.upper() and 'C9' in c10_val.upper())
        d10_ok = bool(d10_val and isinstance(d10_val, str) and
                  'SUM' in d10_val.upper() and 'D4' in d10_val.upper() and 'D9' in d10_val.upper())

        checks_passed = sum([int(a10_ok), int(b10_ok), int(c10_ok), int(d10_ok)])
        if checks_passed == 4:
            print(f"PASS: Component 5 — total row A10:D10 correct (0.10 pts)")
            total_score += 0.10
        elif checks_passed >= 2:
            print(f"PARTIAL: Component 5 — {checks_passed}/4 total row checks passed: "
                  f"A10={a10_ok}(val={repr(a10_val)}), B10={b10_ok}(val={repr(b10_val)}), "
                  f"C10={c10_ok}, D10={d10_ok}")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — total row incorrect: "
                  f"A10={repr(a10_val)}, B10={repr(b10_val)}, "
                  f"C10={repr(c10_val)}, D10={repr(d10_val)}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Thin borders on A3:D10 (0.10 points)
    # All cells in A3:D10 must have thin borders on all 4 sides
    try:
        cells_with_borders = 0
        total_cells = 0
        for row_idx in range(3, 11):  # rows 3 to 10
            for col_idx in range(1, 5):  # cols A to D
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell, MergedCell):
                    continue
                total_cells += 1
                b = cell.border
                has_all_borders = (b.left.style == 'thin' and
                                   b.right.style == 'thin' and
                                   b.top.style == 'thin' and
                                   b.bottom.style == 'thin')
                if has_all_borders:
                    cells_with_borders += 1
                else:
                    print(f"FAIL: Component 6 — {cell.coordinate} missing border: "
                          f"left={b.left.style}, right={b.right.style}, "
                          f"top={b.top.style}, bottom={b.bottom.style}")

        if cells_with_borders == total_cells and total_cells > 0:
            print(f"PASS: Component 6 — all {total_cells} cells in A3:D10 have thin borders (0.10 pts)")
            total_score += 0.10
        elif cells_with_borders >= total_cells * 0.8:
            print(f"PARTIAL: Component 6 — {cells_with_borders}/{total_cells} cells have thin borders (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 — only {cells_with_borders}/{total_cells} cells have thin borders")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
