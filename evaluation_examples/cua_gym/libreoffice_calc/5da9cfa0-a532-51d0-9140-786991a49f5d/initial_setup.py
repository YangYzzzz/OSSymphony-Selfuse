"""
Initial Setup: EEO-1 HR Headcount Report
Task ID: calc_hr_eeo_report_011
Domain: libreoffice_calc

Creates:
- Sheet 'HR Data' with 211 employee records (rows 2-212)
  Headers: Emp ID, Name, Gender, Ethnicity, Department, Job Category, Status
- Sheet 'EEO Summary' (empty, ready for agent to fill)
"""

import openpyxl
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_eeo_report_011'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

random.seed(42)

# Realistic employee name pools
FIRST_NAMES_M = [
    'James', 'Robert', 'Michael', 'William', 'David', 'Richard', 'Joseph',
    'Thomas', 'Charles', 'Christopher', 'Daniel', 'Matthew', 'Anthony',
    'Mark', 'Donald', 'Steven', 'Paul', 'Andrew', 'Joshua', 'Kenneth',
    'Kevin', 'Brian', 'George', 'Timothy', 'Ronald', 'Edward', 'Jason',
    'Jeffrey', 'Ryan', 'Jacob', 'Gary', 'Nicholas', 'Eric', 'Jonathan',
    'Stephen', 'Larry', 'Justin', 'Scott', 'Brandon', 'Benjamin'
]
FIRST_NAMES_F = [
    'Mary', 'Patricia', 'Jennifer', 'Linda', 'Barbara', 'Elizabeth',
    'Susan', 'Jessica', 'Sarah', 'Karen', 'Lisa', 'Nancy', 'Betty',
    'Margaret', 'Sandra', 'Ashley', 'Dorothy', 'Kimberly', 'Emily',
    'Donna', 'Michelle', 'Carol', 'Amanda', 'Melissa', 'Deborah',
    'Stephanie', 'Rebecca', 'Sharon', 'Laura', 'Cynthia', 'Kathleen',
    'Amy', 'Angela', 'Shirley', 'Anna', 'Brenda', 'Pamela', 'Emma',
    'Nicole', 'Helen', 'Samantha', 'Katherine', 'Christine', 'Debra'
]
FIRST_NAMES_NB = [
    'Alex', 'Jordan', 'Taylor', 'Morgan', 'Casey', 'Riley', 'Jamie',
    'Avery', 'Peyton', 'Quinn', 'Skyler', 'Drew', 'Cameron', 'Parker',
    'Sage', 'River', 'Kendall', 'Reese', 'Emerson', 'Finley'
]
LAST_NAMES = [
    'Smith', 'Johnson', 'Williams', 'Brown', 'Jones', 'Garcia', 'Miller',
    'Davis', 'Rodriguez', 'Martinez', 'Hernandez', 'Lopez', 'Gonzalez',
    'Wilson', 'Anderson', 'Thomas', 'Taylor', 'Moore', 'Jackson', 'Martin',
    'Lee', 'Perez', 'Thompson', 'White', 'Harris', 'Sanchez', 'Clark',
    'Ramirez', 'Lewis', 'Robinson', 'Walker', 'Young', 'Allen', 'King',
    'Wright', 'Scott', 'Torres', 'Nguyen', 'Hill', 'Flores', 'Green',
    'Adams', 'Nelson', 'Baker', 'Hall', 'Rivera', 'Campbell', 'Mitchell',
    'Carter', 'Roberts', 'Chen', 'Patel', 'Kim', 'Okonkwo', 'Nakamura'
]

GENDERS = ['Male', 'Female', 'Non-Binary']
ETHNICITIES = ['White', 'Black', 'Hispanic', 'Asian', 'Two or More Races', 'Other']
DEPARTMENTS = [
    'Engineering', 'Marketing', 'Finance', 'Operations', 'Human Resources',
    'Sales', 'Legal', 'Product', 'Customer Success', 'IT'
]
JOB_CATEGORIES = [
    'Officials & Managers', 'Professionals', 'Technicians',
    'Sales Workers', 'Administrative Support', 'Craft Workers',
    'Operatives', 'Service Workers'
]
STATUSES = ['Active', 'Inactive']

# Distribution weights for realistic population
GENDER_WEIGHTS = [0.52, 0.44, 0.04]
ETHNICITY_WEIGHTS = [0.55, 0.13, 0.18, 0.06, 0.05, 0.03]
STATUS_WEIGHTS = [0.85, 0.15]


def generate_employees(n=211):
    employees = []
    for i in range(1, n + 1):
        gender = random.choices(GENDERS, weights=GENDER_WEIGHTS)[0]
        if gender == 'Male':
            first = random.choice(FIRST_NAMES_M)
        elif gender == 'Female':
            first = random.choice(FIRST_NAMES_F)
        else:
            first = random.choice(FIRST_NAMES_NB)
        last = random.choice(LAST_NAMES)
        name = f'{first} {last}'
        ethnicity = random.choices(ETHNICITIES, weights=ETHNICITY_WEIGHTS)[0]
        dept = random.choice(DEPARTMENTS)
        job_cat = random.choice(JOB_CATEGORIES)
        status = random.choices(STATUSES, weights=STATUS_WEIGHTS)[0]
        emp_id = f'E{1000 + i:04d}'
        employees.append([emp_id, name, gender, ethnicity, dept, job_cat, status])
    return employees


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: HR Data ---
    ws_hr = wb.active
    ws_hr.title = 'HR Data'

    headers = ['Emp ID', 'Name', 'Gender', 'Ethnicity', 'Department', 'Job Category', 'Status']
    for col, h in enumerate(headers, 1):
        ws_hr.cell(row=1, column=col, value=h)

    employees = generate_employees(211)
    for r, row_data in enumerate(employees, 2):
        for c, val in enumerate(row_data, 1):
            ws_hr.cell(row=r, column=c, value=val)

    # Set reasonable column widths for readability
    ws_hr.column_dimensions['A'].width = 10
    ws_hr.column_dimensions['B'].width = 22
    ws_hr.column_dimensions['C'].width = 12
    ws_hr.column_dimensions['D'].width = 18
    ws_hr.column_dimensions['E'].width = 20
    ws_hr.column_dimensions['F'].width = 25
    ws_hr.column_dimensions['G'].width = 10

    # --- Sheet 2: EEO Summary (empty — agent must fill) ---
    ws_eeo = wb.create_sheet('EEO Summary')
    # Intentionally empty

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  HR Data rows: {ws_hr.max_row - 1} employee records')
    print(f'  EEO Summary: empty (ready for agent)')

    # Print quick stats for verification
    active_male = sum(1 for e in employees if e[2] == 'Male' and e[6] == 'Active')
    active_female = sum(1 for e in employees if e[2] == 'Female' and e[6] == 'Active')
    active_nb = sum(1 for e in employees if e[2] == 'Non-Binary' and e[6] == 'Active')
    print(f'  Active Males: {active_male}, Active Females: {active_female}, Active Non-Binary: {active_nb}')
    for eth in ['White', 'Black', 'Hispanic', 'Asian', 'Two or More Races', 'Other']:
        m = sum(1 for e in employees if e[2] == 'Male' and e[3] == eth and e[6] == 'Active')
        f = sum(1 for e in employees if e[2] == 'Female' and e[3] == eth and e[6] == 'Active')
        print(f'    {eth}: Male={m}, Female={f}')


create_initial()
