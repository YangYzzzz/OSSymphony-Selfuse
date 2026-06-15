"""
Reward Script: Fill missing categories in source data and refresh pivot table
Task ID: calc_tbl_050
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): All blank category cells in SourceData column B filled correctly
  Component 2 (0.3): PivotReport has no "(blank)"/"(empty)" entries
  Component 3 (0.2): PivotReport shows all 5 expected categories with correct totals
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_050'

# Expected category fill pattern: each group of 4 rows shares the same category
# Row 2-5: Electronics, Row 6-9: Furniture, Row 10-13: Office Supplies,
# Row 14-17: Technology, Row 18-21: Stationery
EXPECTED_CATEGORIES = {
    2: 'Electronics', 3: 'Electronics', 4: 'Electronics', 5: 'Electronics',
    6: 'Furniture', 7: 'Furniture', 8: 'Furniture', 9: 'Furniture',
    10: 'Office Supplies', 11: 'Office Supplies', 12: 'Office Supplies', 13: 'Office Supplies',
    14: 'Technology', 15: 'Technology', 16: 'Technology', 17: 'Technology',
    18: 'Stationery', 19: 'Stationery', 20: 'Stationery', 21: 'Stationery',
}

# Rows that were initially blank (these are the ones the task asks to fill)
INITIALLY_BLANK_ROWS = [3, 4, 5, 7, 8, 9, 11, 12, 13, 15, 16, 17, 19, 20, 21]

# Expected pivot totals per category (sum of Revenue and Quantity from source data)
EXPECTED_PIVOT = {
    'Electronics': {'revenue': 199.44, 'quantity': 404},
    'Furniture': {'revenue': 942.50, 'quantity': 157},
    'Office Supplies': {'revenue': 53.23, 'quantity': 1465},
    'Technology': {'revenue': 313.98, 'quantity': 425},
    'Stationery': {'revenue': 121.49, 'quantity': 345},
}


def persist_app_state(domain):
    """Send Ctrl+S to save any unsaved changes in LibreOffice."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that required sheets exist
    if 'SourceData' not in wb.sheetnames:
        print("CRITICAL: 'SourceData' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws_source = wb['SourceData']

    # ----------------------------------------------------------------
    # Component 1: All initially-blank category cells filled correctly (0.5 points)
    # In the initial file, rows 3-5, 7-9, 11-13, 15-17, 19-21 have blank B column.
    # The task requires filling them with the correct category from surrounding context.
    # ----------------------------------------------------------------
    try:
        filled_count = 0
        correct_count = 0
        for row_num in INITIALLY_BLANK_ROWS:
            cell_val = ws_source.cell(row=row_num, column=2).value
            expected = EXPECTED_CATEGORIES[row_num]
            if cell_val is not None and str(cell_val).strip() != '':
                filled_count += 1
                if str(cell_val).strip() == expected:
                    correct_count += 1
                else:
                    print(f"  MISMATCH: B{row_num} = '{cell_val}', expected '{expected}'")

        total_blank = len(INITIALLY_BLANK_ROWS)  # 15
        if correct_count == total_blank:
            print(f"PASS: Component 1 — All {total_blank} blank cells filled correctly (0.5 pts)")
            total_score += 0.5
        elif correct_count > 0:
            # Partial credit: proportional to how many were filled correctly
            partial = round(0.5 * (correct_count / total_blank), 2)
            print(f"PARTIAL: Component 1 — {correct_count}/{total_blank} cells correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — {filled_count} filled, {correct_count} correct out of {total_blank}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ----------------------------------------------------------------
    # Component 2: PivotReport has no "(blank)" or "(empty)" entries (0.3 points)
    # In the initial file, the pivot table shows "(empty)" as a category.
    # After filling and refreshing, this should disappear.
    # ----------------------------------------------------------------
    try:
        has_pivot_sheet = 'PivotReport' in wb.sheetnames
        if not has_pivot_sheet:
            print("FAIL: Component 2 — 'PivotReport' sheet not found")
        else:
            ws_pivot = wb['PivotReport']
            found_blank_entry = False
            for row in ws_pivot.iter_rows(min_row=1, max_row=ws_pivot.max_row, min_col=1, max_col=ws_pivot.max_column):
                for cell in row:
                    if cell.value is not None:
                        val_str = str(cell.value).strip().lower()
                        if val_str in ('(blank)', '(empty)', 'blank', 'empty'):
                            found_blank_entry = True
                            print(f"  Found blank entry at {cell.coordinate}: '{cell.value}'")

            if not found_blank_entry:
                print(f"PASS: Component 2 — No blank/empty entries in PivotReport (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — PivotReport still contains blank/empty entries")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ----------------------------------------------------------------
    # Component 3: PivotReport revenue totals match expected values (0.2 points)
    # After filling all categories and refreshing, each category's revenue sum
    # should reflect ALL rows, not just the first row of each group.
    # Initial pivot has wrong totals (e.g., Electronics=29.99 instead of 199.44).
    # We load with data_only=True to read cached computed values, then check
    # that at least 3 out of 5 category revenue totals match expected values.
    # ----------------------------------------------------------------
    try:
        if 'PivotReport' not in wb.sheetnames:
            print("FAIL: Component 3 — 'PivotReport' sheet not found")
        else:
            # Reload with data_only=True to get cached values from pivot
            import openpyxl as openpyxl2
            wb_data = openpyxl2.load_workbook(file_path, data_only=True)
            ws_pv = wb_data['PivotReport']

            # Scan pivot for category revenue values
            # The pivot could be in various formats. We look for category names
            # and associated revenue numbers nearby.
            matched_cats = 0
            for row_idx in range(1, ws_pv.max_row + 1):
                a_val = ws_pv.cell(row=row_idx, column=1).value
                if a_val is not None and str(a_val).strip() in EXPECTED_PIVOT:
                    cat_name = str(a_val).strip()
                    expected_rev = EXPECTED_PIVOT[cat_name]['revenue']
                    # Check column B (or nearby cells) for the revenue value
                    # The golden pivot format: A=category, B=Sum of Revenue, C=Sum of Quantity
                    b_val = ws_pv.cell(row=row_idx, column=2).value
                    # Also check rows below (initial pivot has Revenue on next row)
                    found_match = False
                    for check_row in [row_idx, row_idx + 1, row_idx + 2]:
                        for check_col in [2, 3]:
                            cv = ws_pv.cell(row=check_row, column=check_col).value
                            if cv is not None:
                                try:
                                    if abs(float(cv) - expected_rev) < 0.1:
                                        found_match = True
                                        break
                                except (ValueError, TypeError):
                                    pass
                        if found_match:
                            break
                    if found_match:
                        matched_cats += 1
                        print(f"  Component 3: {cat_name} revenue matches expected {expected_rev}")
                    else:
                        print(f"  Component 3: {cat_name} revenue does NOT match expected {expected_rev}")

            if matched_cats >= 4:
                print(f"PASS: Component 3 — {matched_cats}/5 category revenues correct (0.2 pts)")
                total_score += 0.2
            elif matched_cats >= 2:
                partial = round(0.2 * (matched_cats / 5), 2)
                print(f"PARTIAL: Component 3 — {matched_cats}/5 category revenues correct ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 3 — Only {matched_cats}/5 category revenues match")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
