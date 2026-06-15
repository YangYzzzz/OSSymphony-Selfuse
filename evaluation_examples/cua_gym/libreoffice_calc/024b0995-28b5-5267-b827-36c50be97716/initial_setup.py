"""
Initial Setup: Fill missing categories in source data and refresh pivot table
Task ID: calc_tbl_050
Domain: libreoffice_calc

Creates a spreadsheet with:
- SourceData sheet: sales data where column B (Category) has blank cells in grouped sections
- PivotReport sheet: a DataPilot pivot table summarizing by category, showing '(blank)'
  for the empty category cells

The pivot table is created via LibreOffice's UNO DataPilot API.
"""

import os
import shlex
import subprocess
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_050'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def run_cmd(command: str, timeout: int = 120):
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    result = subprocess.run(
        command, shell=True, capture_output=True, text=True,
        timeout=timeout, env=env
    )
    print(f"CMD: {command}")
    if result.stdout.strip():
        print(f"OUT: {result.stdout[:500]}")
    if result.stderr.strip():
        print(f"ERR: {result.stderr[:500]}")
    return result


def create_source_file():
    """Create xlsx with source data using openpyxl (no pivot yet)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'SourceData'

    # Headers
    headers = ['ID', 'Category', 'Product', 'Revenue', 'Quantity']
    white_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data with blank Category cells - first row of group has category, rest blank
    data = [
        [1001, 'Electronics',     'Wireless Mouse',        29.99,  145],
        [1002, '',                'USB-C Hub',              45.50,   87],
        [1003, '',                'Mechanical Keyboard',    89.00,   62],
        [1004, '',                'Monitor Stand',          34.95,  110],
        [1005, 'Furniture',       'Standing Desk',         349.00,   28],
        [1006, '',                'Ergonomic Chair',       275.00,   41],
        [1007, '',                'Filing Cabinet',        129.50,   55],
        [1008, '',                'Bookshelf Unit',        189.00,   33],
        [1009, 'Office Supplies', 'Premium Notebooks',      12.99, 320],
        [1010, '',                'Ballpoint Pen Set',       8.50, 450],
        [1011, '',                'Desk Organizer',         24.99, 175],
        [1012, '',                'Sticky Notes Pack',       6.75, 520],
        [1013, 'Technology',      'Webcam HD 1080p',        59.99,  95],
        [1014, '',                'Noise-Cancel Headset',  149.00,   48],
        [1015, '',                'Portable SSD 1TB',       79.99,   72],
        [1016, '',                'Wireless Charger',       25.00, 210],
        [1017, 'Stationery',      'Fountain Pen',           42.00,  65],
        [1018, '',                'Watercolor Set',         35.50,   88],
        [1019, '',                'Sketch Pad A3',          15.99, 140],
        [1020, '',                'Calligraphy Kit',        28.00,  52],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    for r in range(2, 22):
        ws.cell(row=r, column=4).number_format = '$#,##0.00'

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12

    # Create empty PivotReport sheet (pivot table added via LO macro)
    wb.create_sheet('PivotReport')

    wb.save(OUTPUT)
    print(f'Source file saved: {OUTPUT}')


def add_pivot_table():
    """Use LibreOffice headless + Basic macro to create a DataPilot pivot table."""

    # Kill any running LO
    run_cmd('pkill -9 -f soffice 2>/dev/null; sleep 3')

    # Write a Basic macro
    macro_text = r"""Sub CreatePivotTable(sFilePath As String)
    Dim sURL As String
    sURL = ConvertToURL(sFilePath)

    Dim oDoc As Object
    oDoc = StarDesktop.loadComponentFromURL(sURL, "_blank", 0, Array())

    If IsNull(oDoc) Then
        Print "ERROR: Could not open document"
        Exit Sub
    End If

    Dim oSheets As Object
    oSheets = oDoc.Sheets
    Dim oSource As Object
    oSource = oSheets.getByName("SourceData")
    Dim oPivot As Object
    oPivot = oSheets.getByName("PivotReport")

    ' Source range A1:E21
    Dim oRange As Object
    oRange = oSource.getCellRangeByName("A1:E21")

    ' DataPilot on PivotReport
    Dim oDPT As Object
    oDPT = oPivot.getDataPilotTables()
    Dim oTarget As Object
    oTarget = oPivot.getCellByPosition(0, 0).getCellAddress()

    Dim oDesc As Object
    oDesc = oDPT.createDataPilotDescriptor()
    oDesc.setSourceRange(oRange.getRangeAddress())

    Dim oFields As Object
    oFields = oDesc.getDataPilotFields()
    Dim i As Integer
    For i = 0 To oFields.Count - 1
        Dim oField As Object
        oField = oFields.getByIndex(i)
        Select Case oField.Name
            Case "Category"
                oField.Orientation = com.sun.star.sheet.DataPilotFieldOrientation.ROW
            Case "Revenue"
                oField.Orientation = com.sun.star.sheet.DataPilotFieldOrientation.DATA
                oField.Function = com.sun.star.sheet.GeneralFunction.SUM
            Case "Quantity"
                oField.Orientation = com.sun.star.sheet.DataPilotFieldOrientation.DATA
                oField.Function = com.sun.star.sheet.GeneralFunction.SUM
        End Select
    Next i

    oDPT.insertNewByName("SalesPivot", oTarget, oDesc)

    oDoc.store()
    oDoc.close(True)
End Sub"""

    # Install macro in user space
    basic_dir = os.path.expanduser('~/.config/libreoffice/4/user/basic/Standard')
    os.makedirs(basic_dir, exist_ok=True)

    xba_path = os.path.join(basic_dir, 'PivotMod.xba')
    xba_content = f'''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="PivotMod" script:language="StarBasic">{macro_text}</script:module>'''

    with open(xba_path, 'w') as f:
        f.write(xba_content)

    # Ensure script.xlb references the module
    xlb_path = os.path.join(basic_dir, 'script.xlb')
    xlb_content = '''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE library:library PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "library.dtd">
<library:library xmlns:library="http://openoffice.org/2000/library" library:name="Standard" library:readonly="false" library:passwordprotected="false">
 <library:element library:name="PivotMod"/>
</library:library>'''
    with open(xlb_path, 'w') as f:
        f.write(xlb_content)

    print('Basic macro installed')

    # Run macro headless
    result = run_cmd(
        f'soffice --headless --invisible '
        f'"macro:///Standard.PivotMod.CreatePivotTable({OUTPUT})"',
        timeout=60
    )
    time.sleep(5)
    run_cmd('pkill -9 -f soffice 2>/dev/null; sleep 2')

    return verify_pivot()


def verify_pivot():
    """Check if PivotReport has data."""
    import openpyxl
    wb = openpyxl.load_workbook(OUTPUT)
    ws = wb['PivotReport']
    has_data = ws['A1'].value is not None
    print(f'PivotReport has data: {has_data}')
    wb.close()
    return has_data


def create_manual_pivot():
    """
    Fallback: Create a manual pivot-style summary on PivotReport sheet.
    Shows (blank) entry for rows with empty category, mimicking DataPilot output.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.load_workbook(OUTPUT)

    if 'PivotReport' in wb.sheetnames:
        del wb['PivotReport']
    ws = wb.create_sheet('PivotReport')

    bold = Font(bold=True)
    header_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")

    # Pivot headers
    ws['A1'] = 'Category'
    ws['B1'] = 'Sum of Revenue'
    ws['C1'] = 'Sum of Quantity'
    for c in ['A1', 'B1', 'C1']:
        ws[c].font = bold
        ws[c].fill = header_fill

    # Compute from SourceData
    src = wb['SourceData']
    cats = {}
    for r in range(2, 22):
        cat = src.cell(row=r, column=2).value
        if not cat or str(cat).strip() == '':
            cat = '(blank)'
        rev = src.cell(row=r, column=4).value or 0
        qty = src.cell(row=r, column=5).value or 0
        if cat not in cats:
            cats[cat] = [0.0, 0]
        cats[cat][0] += float(rev)
        cats[cat][1] += int(qty)

    # Sort: (blank) first, then alphabetical
    sorted_keys = sorted([k for k in cats if k != '(blank)'])
    if '(blank)' in cats:
        sorted_keys.insert(0, '(blank)')

    row = 2
    t_rev, t_qty = 0.0, 0
    for cat in sorted_keys:
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=round(cats[cat][0], 2))
        ws.cell(row=row, column=2).number_format = '$#,##0.00'
        ws.cell(row=row, column=3, value=cats[cat][1])
        t_rev += cats[cat][0]
        t_qty += cats[cat][1]
        row += 1

    # Grand Total
    ws.cell(row=row, column=1, value='Grand Total')
    ws.cell(row=row, column=1).font = bold
    ws.cell(row=row, column=2, value=round(t_rev, 2))
    ws.cell(row=row, column=2).number_format = '$#,##0.00'
    ws.cell(row=row, column=2).font = bold
    ws.cell(row=row, column=3, value=t_qty)
    ws.cell(row=row, column=3).font = bold

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18

    wb.save(OUTPUT)
    print('Manual pivot layout created with (blank) entries')


def create_initial():
    # Step 1: Create source xlsx
    create_source_file()

    # Step 2: Try to add real pivot table via LibreOffice macro
    pivot_ok = add_pivot_table()

    # Step 3: Fallback to manual pivot if macro failed
    if not pivot_ok:
        print('Macro approach failed, using manual pivot layout')
        create_manual_pivot()

    # Step 4: Open file in GUI
    run_cmd('pkill -9 -f soffice 2>/dev/null; sleep 2')
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
