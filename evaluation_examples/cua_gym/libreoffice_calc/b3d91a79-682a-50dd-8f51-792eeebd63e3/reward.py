"""
Reward Script: CRM data cleanup - convert text-with-NBSP to numeric values and add totals row
Task ID: calc_gen_data_cleanup_012
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Revenue column D2:D101 converted to numeric (0.30 pts)
  Component 2: Units Sold E2:E101 and Discount F2:F101 converted to numeric (0.20 pts)
  Component 3: Correct number formats applied to D, E, F columns (0.20 pts)
  Component 4: Totals row 102 with TOTALS label and SUM/AVERAGE formulas (0.30 pts)
  Total: 1.0

Notes:
  - Initial file has D/E/F as text strings prefixed with non-breaking space (CHAR(160), \\xa0)
  - Golden file has real numeric values with currency/integer/percentage formats
  - Row 102 should have: A102='TOTALS', D102=SUM(D2:D101), E102=SUM(E2:E101), F102=AVERAGE(F2:F101)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_data_cleanup_012'
SHEET_NAME = 'CRMExport'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Found: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Component 1: Revenue column D2:D101 converted to real numeric values (0.30 pts)
    # In the initial file, D2:D101 contain text strings like '\xa048500' (NBSP prefix)
    # Task requires converting them to real numbers (int or float)
    try:
        numeric_count_d = 0
        text_count_d = 0
        for row in range(2, 102):
            val = ws.cell(row=row, column=4).value  # Column D
            if val is None:
                text_count_d += 1
            elif isinstance(val, (int, float)):
                numeric_count_d += 1
            elif isinstance(val, str):
                # Check if it still has the NBSP prefix or is non-numeric
                text_count_d += 1
            else:
                text_count_d += 1

        if numeric_count_d == 100:
            print(f"PASS: Component 1 — Revenue column D2:D101 all numeric ({numeric_count_d}/100 cells are numeric) (0.30 pts)")
            total_score += 0.30
        elif numeric_count_d >= 50:
            print(f"PARTIAL: Component 1 — Revenue column D2:D101 partially numeric ({numeric_count_d}/100) — no partial credit awarded")
            print(f"FAIL: Component 1 — expected 100/100 numeric, found {numeric_count_d}/100")
        else:
            print(f"FAIL: Component 1 — Revenue column D2:D101 still text ({text_count_d}/100 are non-numeric)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Units Sold E2:E101 and Discount F2:F101 converted to numeric (0.20 pts)
    # Both columns must be fully numeric
    try:
        numeric_count_e = 0
        numeric_count_f = 0
        for row in range(2, 102):
            val_e = ws.cell(row=row, column=5).value  # Column E
            val_f = ws.cell(row=row, column=6).value  # Column F
            if isinstance(val_e, (int, float)):
                numeric_count_e += 1
            if isinstance(val_f, (int, float)):
                numeric_count_f += 1

        if numeric_count_e == 100 and numeric_count_f == 100:
            print(f"PASS: Component 2 — Units Sold E ({numeric_count_e}/100) and Discount F ({numeric_count_f}/100) all numeric (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 — Units E={numeric_count_e}/100 numeric, Discount F={numeric_count_f}/100 numeric (expected 100/100 each)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Correct number formats applied to D, E, F columns (0.20 pts)
    # D: currency format ($#,##0.00), E: integer format (0), F: percentage format (0.00%)
    # Check on a sample of cells (rows 2, 3, 4) and the totals row if present
    try:
        d_fmt_ok = 0
        e_fmt_ok = 0
        f_fmt_ok = 0
        sample_rows = list(range(2, 12))  # Check first 10 data rows

        for row in sample_rows:
            nf_d = ws.cell(row=row, column=4).number_format
            nf_e = ws.cell(row=row, column=5).number_format
            nf_f = ws.cell(row=row, column=6).number_format

            # Revenue should have currency format (contains $ or #,##0)
            if '$' in nf_d or ('#,##0' in nf_d):
                d_fmt_ok += 1

            # Units sold should be integer format
            if nf_e in ('0', '0.0', '#,##0', 'General') or nf_e == '0':
                # Accept '0' or similar integer formats
                if nf_e == '0':
                    e_fmt_ok += 1

            # Discount should be percentage format
            if '%' in nf_f:
                f_fmt_ok += 1

        # For E, also check all rows since format '0' is specific
        e_fmt_ok_all = 0
        for row in sample_rows:
            nf_e = ws.cell(row=row, column=5).number_format
            if nf_e == '0':
                e_fmt_ok_all += 1

        formats_pass = (d_fmt_ok == len(sample_rows) and
                        e_fmt_ok_all == len(sample_rows) and
                        f_fmt_ok == len(sample_rows))

        if formats_pass:
            print(f"PASS: Component 3 — D currency ({d_fmt_ok}/{len(sample_rows)}), E integer ({e_fmt_ok_all}/{len(sample_rows)}), F percentage ({f_fmt_ok}/{len(sample_rows)}) (0.20 pts)")
            total_score += 0.20
        else:
            # Partial: check what formats are actually applied
            sample_d_nf = ws.cell(row=2, column=4).number_format
            sample_e_nf = ws.cell(row=2, column=5).number_format
            sample_f_nf = ws.cell(row=2, column=6).number_format
            print(f"FAIL: Component 3 — Format check failed:")
            print(f"  D2 format='{sample_d_nf}' (need '$#,##0.00' style) — {d_fmt_ok}/{len(sample_rows)} ok")
            print(f"  E2 format='{sample_e_nf}' (need '0') — {e_fmt_ok_all}/{len(sample_rows)} ok")
            print(f"  F2 format='{sample_f_nf}' (need '0.00%' style) — {f_fmt_ok}/{len(sample_rows)} ok")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Totals row 102 with TOTALS label and SUM/AVERAGE formulas (0.30 pts)
    # A102 = 'TOTALS'
    # D102 = SUM formula over D2:D101
    # E102 = SUM formula over E2:E101
    # F102 = AVERAGE formula over F2:F101
    try:
        a102 = ws.cell(row=102, column=1).value
        d102 = ws.cell(row=102, column=4).value
        e102 = ws.cell(row=102, column=5).value
        f102 = ws.cell(row=102, column=6).value

        # Check A102 label: must contain 'TOTAL'
        a102_ok = (isinstance(a102, str) and 'TOTAL' in a102.upper())

        # Check D102: SUM formula referencing D2:D101 (or numeric value matching expected sum)
        d102_is_sum_formula = (isinstance(d102, str) and 'SUM' in d102.upper() and 'D2' in d102.upper())
        d102_is_numeric_sum = (isinstance(d102, (int, float)) and abs(d102 - 5701100.0) <= 1.0)
        d102_ok = d102_is_sum_formula or d102_is_numeric_sum

        # Check E102: SUM formula referencing E2:E101 (or numeric value matching expected sum)
        e102_is_sum_formula = (isinstance(e102, str) and 'SUM' in e102.upper() and 'E2' in e102.upper())
        e102_is_numeric_sum = (isinstance(e102, (int, float)) and abs(e102 - 1776.0) <= 1.0)
        e102_ok = e102_is_sum_formula or e102_is_numeric_sum

        # Check F102: AVERAGE formula referencing F2:F101 (or numeric value matching expected average)
        f102_is_avg_formula = (isinstance(f102, str) and 'AVERAGE' in f102.upper() and 'F2' in f102.upper())
        f102_is_numeric_avg = (isinstance(f102, (int, float)) and abs(f102 - 0.1466) <= 0.01)
        f102_ok = f102_is_avg_formula or f102_is_numeric_avg

        all_ok = a102_ok and d102_ok and e102_ok and f102_ok

        if all_ok:
            print(f"PASS: Component 4 — Totals row 102: A102={repr(a102)}, D102={repr(d102)}, E102={repr(e102)}, F102={repr(f102)} (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 4 — Totals row 102 incomplete:")
            print(f"  A102={repr(a102)} (need 'TOTALS') — {'OK' if a102_ok else 'FAIL'}")
            print(f"  D102={repr(d102)} (need SUM(D2:D101)) — {'OK' if d102_ok else 'FAIL'}")
            print(f"  E102={repr(e102)} (need SUM(E2:E101)) — {'OK' if e102_ok else 'FAIL'}")
            print(f"  F102={repr(f102)} (need AVERAGE(F2:F101)) — {'OK' if f102_ok else 'FAIL'}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
