"""
Initial Setup: CRM Export with text-formatted numeric columns (non-breaking space prefix)
Task ID: calc_gen_data_cleanup_012
Domain: libreoffice_calc

Creates a CRMExport sheet with 100 rows where Revenue, Units Sold, and Discount %
columns are stored as text strings with a non-breaking space (chr(160)) prefix,
simulating a CRM export where numbers were exported as text.
"""

import openpyxl
from openpyxl.styles import Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_data_cleanup_012'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

# Non-breaking space character (CHAR(160))
NBSP = '\u00a0'

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'CRMExport'

    # --- Headers ---
    headers = ['Opp ID', 'Account', 'Rep', 'Revenue', 'Units Sold', 'Discount %', 'Quarter']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic CRM data
    accounts = [
        'Nexus Systems', 'Apex Dynamics', 'Blue Ridge Corp', 'Summit Tech',
        'Vantage Group', 'Orion Enterprises', 'Pinnacle Solutions', 'Harbor Networks',
        'Crestview Inc', 'Sterling Capital', 'Atlas Global', 'Meridian Partners',
        'Cobalt Industries', 'Phoenix Digital', 'Granite Holdings',
        'Tidewater Corp', 'Silverline Tech', 'Cascade Systems', 'Vertex Analytics',
        'Ironbridge Co'
    ]
    reps = [
        'Sarah Chen', 'Marcus Johnson', 'Diana Torres', 'Kevin Park',
        'Amanda Reyes', 'Brian Walsh', 'Lisa Nguyen', 'Carlos Rivera',
        'Patricia Holt', 'Derek Simmons'
    ]
    quarters = ['Q1-2025', 'Q2-2025', 'Q3-2025', 'Q4-2025']

    # Revenue values (realistic deal sizes in dollars)
    revenues = [
        48500, 72300, 31200, 95400, 18700, 63800, 47200, 82100, 29500, 54300,
        71000, 38600, 91200, 25400, 67800, 43100, 58700, 34900, 79300, 52600,
        41800, 86500, 27300, 61400, 74200, 33700, 49900, 88300, 21600, 65100,
        57400, 93800, 36200, 45700, 70500, 28900, 83600, 50200, 39400, 76800,
        62300, 44500, 97100, 31800, 55200, 69700, 42400, 87900, 24700, 58500,
        73600, 35100, 48300, 81700, 26500, 64200, 51800, 92400, 37600, 46900,
        68400, 53700, 85200, 29800, 61900, 75300, 40600, 89700, 23400, 57100,
        66800, 45200, 78500, 32300, 53900, 72700, 38100, 91500, 27600, 60400,
        49800, 84100, 36700, 54600, 77200, 43800, 69300, 25900, 88600, 51400,
        70900, 34500, 59200, 82800, 46300, 63700, 28100, 76500, 41200, 94300
    ]

    # Units sold (integer values)
    units = [
        12, 23, 8, 31, 5, 19, 14, 27, 9, 16,
        22, 11, 28, 7, 21, 13, 18, 10, 25, 15,
        20, 26, 8, 19, 24, 10, 15, 29, 6, 20,
        17, 30, 11, 14, 22, 9, 27, 16, 12, 24,
        19, 13, 32, 10, 17, 21, 13, 28, 7, 18,
        23, 11, 15, 26, 8, 20, 16, 30, 12, 14,
        21, 17, 27, 9, 19, 24, 13, 28, 7, 18,
        21, 14, 25, 10, 17, 23, 12, 29, 8, 19,
        15, 27, 11, 17, 24, 14, 22, 8, 28, 16,
        22, 11, 18, 26, 14, 20, 9, 24, 13, 30
    ]

    # Discount percentages (stored as decimal, e.g., 0.10 for 10%)
    discounts = [
        0.10, 0.15, 0.05, 0.20, 0.08, 0.12, 0.18, 0.07, 0.25, 0.10,
        0.15, 0.05, 0.22, 0.08, 0.12, 0.17, 0.10, 0.06, 0.20, 0.14,
        0.09, 0.18, 0.05, 0.13, 0.21, 0.07, 0.11, 0.25, 0.04, 0.16,
        0.12, 0.23, 0.08, 0.10, 0.19, 0.06, 0.24, 0.13, 0.09, 0.17,
        0.15, 0.11, 0.28, 0.07, 0.14, 0.20, 0.09, 0.26, 0.05, 0.18,
        0.22, 0.08, 0.12, 0.19, 0.06, 0.15, 0.11, 0.24, 0.10, 0.13,
        0.17, 0.14, 0.22, 0.07, 0.16, 0.21, 0.09, 0.27, 0.05, 0.15,
        0.19, 0.11, 0.23, 0.08, 0.14, 0.20, 0.10, 0.28, 0.06, 0.17,
        0.13, 0.25, 0.09, 0.16, 0.22, 0.12, 0.18, 0.07, 0.26, 0.14,
        0.20, 0.10, 0.15, 0.24, 0.11, 0.19, 0.08, 0.23, 0.13, 0.30
    ]

    left_align = Alignment(horizontal='left')

    for i in range(100):
        row = i + 2
        opp_id = f'OPP-{2025001 + i}'
        account = accounts[i % len(accounts)]
        rep = reps[i % len(reps)]
        quarter = quarters[i % len(quarters)]

        # Revenue as text with NBSP prefix (left-aligned to confirm text)
        revenue_text = f'{NBSP}{revenues[i]}'
        # Units Sold as text with NBSP prefix
        units_text = f'{NBSP}{units[i]}'
        # Discount as text with NBSP prefix (stored as percentage representation e.g. "0.10")
        discount_text = f'{NBSP}{discounts[i]}'

        ws.cell(row=row, column=1, value=opp_id)
        ws.cell(row=row, column=2, value=account)
        ws.cell(row=row, column=3, value=rep)

        # Write text values for D, E, F and force left alignment
        cell_d = ws.cell(row=row, column=4, value=revenue_text)
        cell_d.alignment = left_align

        cell_e = ws.cell(row=row, column=5, value=units_text)
        cell_e.alignment = left_align

        cell_f = ws.cell(row=row, column=6, value=discount_text)
        cell_f.alignment = left_align

        ws.cell(row=row, column=7, value=quarter)

    # Row 102 intentionally left empty

    # Set some column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: CRMExport')
    print(f'  Headers in row 1, data in rows 2-101 (100 rows)')
    print(f'  Columns D, E, F: text with non-breaking space prefix')
    print(f'  Row 102: empty')

create_initial()
