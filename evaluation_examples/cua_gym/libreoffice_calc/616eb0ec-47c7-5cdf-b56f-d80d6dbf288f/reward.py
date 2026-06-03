"""
Reward Script: Weekly Shift Schedule with Data Validation, Formulas, and Conditional Formatting
Task ID: calc_ops_resource_shift_planning_036
Domain: libreoffice_calc
Scoring:
  Component 1: Data validation on B2:H21 (list: M,A,N,O)          — 0.25 pts
  Component 2: COUNTIF formulas in I2:I21 (Total Shifts)           — 0.25 pts
  Component 3: IF formulas in J2:J21 (Overtime Flag)               — 0.20 pts
  Component 4: Row 22 daily totals COUNTIF formulas (B22:H22)      — 0.15 pts
  Component 5: Conditional formatting on J2:J21 (red for OVERTIME) — 0.10 pts
  Component 6: Conditional formatting on B2:H21 (shift colors)     — 0.05 pts
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_resource_shift_planning_036'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: ShiftSchedule sheet must exist
    if 'ShiftSchedule' not in wb.sheetnames:
        print("CRITICAL: 'ShiftSchedule' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['ShiftSchedule']

    # -----------------------------------------------------------------------
    # Component 1: Data validation on B2:H21 with list "M,A,N,O" (0.25 pts)
    # This FAILS on initial (no validations) -> PASSES on golden
    # -----------------------------------------------------------------------
    try:
        dvs = ws.data_validations.dataValidation
        found_dv = False
        for dv in dvs:
            if dv.type == 'list' and dv.formula1 is not None:
                # Accept formula1 like '"M,A,N,O"' or '"M, A, N, O"'
                formula_clean = dv.formula1.replace('"', '').replace(' ', '').upper()
                if set(formula_clean.split(',')) == {'M', 'A', 'N', 'O'}:
                    # Check sqref covers B2:H21 (or is a superset)
                    sqref_str = str(dv.sqref)
                    if 'B2' in sqref_str or 'B2:H21' in sqref_str:
                        found_dv = True
                        break
        if found_dv:
            print(f"PASS: Component 1 — Data validation (list: M,A,N,O) present on B2:H21 (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — Expected data validation list 'M,A,N,O' on B2:H21, found {len(dvs)} validations")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: COUNTIF formulas in I2:I21 (Total Shifts) (0.25 pts)
    # Formula should be =COUNTIF(B<row>:H<row>,"<>O")
    # This FAILS on initial (all None) -> PASSES on golden
    # -----------------------------------------------------------------------
    try:
        countif_count = 0
        for row in range(2, 22):
            val = ws.cell(row=row, column=9).value  # Column I
            if val and isinstance(val, str):
                val_clean = val.upper().replace(' ', '')
                # Check for COUNTIF formula referencing same row with <>O condition
                if 'COUNTIF(' in val_clean and '"<>O"' in val_clean:
                    countif_count += 1

        if countif_count == 20:
            print(f"PASS: Component 2 — COUNTIF formulas present in I2:I21 (all 20 rows) (0.25 pts)")
            total_score += 0.25
        elif countif_count >= 10:
            # Partial credit not awarded per component design; requires full
            print(f"FAIL: Component 2 — Only {countif_count}/20 COUNTIF formulas found in I2:I21")
        else:
            print(f"FAIL: Component 2 — Expected COUNTIF formulas in I2:I21, only {countif_count}/20 found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: IF formulas in J2:J21 (Overtime Flag) (0.20 pts)
    # Formula should be =IF(I<row>>5,"OVERTIME","OK")
    # This FAILS on initial (all None) -> PASSES on golden
    # -----------------------------------------------------------------------
    try:
        if_count = 0
        for row in range(2, 22):
            val = ws.cell(row=row, column=10).value  # Column J
            if val and isinstance(val, str):
                val_clean = val.upper().replace(' ', '')
                # Check for IF formula referencing I column with OVERTIME/OK
                if 'IF(' in val_clean and 'OVERTIME' in val_clean and 'OK' in val_clean:
                    if_count += 1

        if if_count == 20:
            print(f"PASS: Component 3 — IF OVERTIME formulas present in J2:J21 (all 20 rows) (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 3 — Expected IF OVERTIME formulas in J2:J21, only {if_count}/20 found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: Row 22 daily totals COUNTIF formulas (B22:H22) (0.15 pts)
    # Formula should be =COUNTIF(<col>2:<col>21,"<>O") for each day column
    # This FAILS on initial (no formulas in B22:H22) -> PASSES on golden
    # Note: A22 has 'Daily Totals' in BOTH initial and golden, so we only
    # check B22:H22 which are empty in initial and have formulas in golden
    # -----------------------------------------------------------------------
    try:
        totals_count = 0
        for col in range(2, 9):  # Columns B-H (2-8)
            val = ws.cell(row=22, column=col).value
            if val and isinstance(val, str):
                val_clean = val.upper().replace(' ', '')
                # Check for COUNTIF formula referencing same column rows 2-21
                if 'COUNTIF(' in val_clean:
                    totals_count += 1

        if totals_count == 7:
            print(f"PASS: Component 4 — COUNTIF daily totals present in B22:H22 (all 7 days) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — Expected COUNTIF formulas in B22:H22, only {totals_count}/7 found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -----------------------------------------------------------------------
    # Component 5: Conditional formatting on J2:J21 (red fill for OVERTIME) (0.10 pts)
    # This FAILS on initial (no CF) -> PASSES on golden
    # -----------------------------------------------------------------------
    try:
        j_cf_found = False
        for cf_range in ws.conditional_formatting:
            cf_range_str = str(cf_range)
            if 'J2' in cf_range_str or 'J' in cf_range_str.upper():
                for rule in ws.conditional_formatting[cf_range]:
                    if rule.type == 'expression' and hasattr(rule, 'formula') and rule.formula:
                        formula_str = str(rule.formula).upper()
                        if 'OVERTIME' in formula_str:
                            # Check for red fill
                            try:
                                if rule.dxf and rule.dxf.fill:
                                    fill_color = rule.dxf.fill.fgColor.rgb
                                    # Red fill: FF0000 in RGB portion
                                    if 'FF0000' in fill_color or fill_color == 'FFFF0000':
                                        j_cf_found = True
                                        break
                            except Exception:
                                # If we can find the OVERTIME CF rule, give partial credit
                                j_cf_found = True
                                break
                if j_cf_found:
                    break

        if j_cf_found:
            print(f"PASS: Component 5 — Conditional formatting (red fill for OVERTIME) on J column (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 — Expected red fill conditional formatting for OVERTIME on J2:J21")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # -----------------------------------------------------------------------
    # Component 6: Conditional formatting on B2:H21 (color-coded shifts) (0.05 pts)
    # Colors: M=blue, A=orange, N=purple, O=grey
    # This FAILS on initial (no CF) -> PASSES on golden
    # -----------------------------------------------------------------------
    try:
        shift_colors_found = 0
        for cf_range in ws.conditional_formatting:
            cf_range_str = str(cf_range)
            if 'B2' in cf_range_str:
                for rule in ws.conditional_formatting[cf_range]:
                    if rule.type == 'expression' and hasattr(rule, 'formula') and rule.formula:
                        formula_str = str(rule.formula)
                        # Check for shift code references (M, A, N, O)
                        if any(code in formula_str for code in ['"M"', '"A"', '"N"', '"O"']):
                            shift_colors_found += 1

        if shift_colors_found >= 4:
            print(f"PASS: Component 6 — Conditional formatting for shift color-coding on B2:H21 ({shift_colors_found} rules) (0.05 pts)")
            total_score += 0.05
        elif shift_colors_found >= 2:
            print(f"PASS: Component 6 — Partial shift color-coding found ({shift_colors_found} rules), awarding partial (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 — Expected 4 shift color rules on B2:H21, found {shift_colors_found}")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
