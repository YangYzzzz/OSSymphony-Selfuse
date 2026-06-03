"""
Initial Setup: Warehouse Team Weekly Shift Schedule
Task ID: calc_ops_resource_shift_planning_036
Domain: libreoffice_calc

Creates the pre-task state:
- Sheet 'ShiftSchedule' with:
  - A1: week commencing date
  - B1:H1: day headers (Mon-Sun)
  - A2:A21: 20 staff member names
  - B2:H21: EMPTY shift assignment grid
  - I1: 'Total Shifts' header, I2:I21 empty
  - J1: 'Overtime Flag' header, J2:J21 empty
  - Row 22: EMPTY (no formulas)
  - NO data validation, NO conditional formatting, NO formulas
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import date

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_resource_shift_planning_036'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ShiftSchedule'

    # --- A1: Week commencing date ---
    ws['A1'] = date(2025, 3, 3)
    ws['A1'].number_format = 'yyyy-mm-dd'
    ws['A1'].font = Font(bold=True, name='Calibri', size=11)

    # --- B1:H1: Day names ---
    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    header_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    header_font = Font(bold=True, name='Calibri', size=11)
    header_align = Alignment(horizontal='center', vertical='center')

    for col, day in enumerate(days, 2):
        cell = ws.cell(row=1, column=col, value=day)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- I1: Total Shifts header ---
    ws['I1'] = 'Total Shifts'
    ws['I1'].font = Font(bold=True, name='Calibri', size=11)
    ws['I1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['I1'].fill = PatternFill(start_color='FFFFF2CC', end_color='FFFFF2CC', fill_type='solid')

    # --- J1: Overtime Flag header ---
    ws['J1'] = 'Overtime Flag'
    ws['J1'].font = Font(bold=True, name='Calibri', size=11)
    ws['J1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['J1'].fill = PatternFill(start_color='FFFFF2CC', end_color='FFFFF2CC', fill_type='solid')

    # --- A2:A21: 20 staff member names ---
    staff_names = [
        'Sarah Chen',
        'Marcus Johnson',
        'Emily Rodriguez',
        'David Kim',
        'Priya Patel',
        'James O\'Brien',
        'Natasha Williams',
        'Carlos Mendez',
        'Aisha Thompson',
        'Ryan Murphy',
        'Linda Zhang',
        'Kevin Okafor',
        'Sophie Turner',
        'Mohammed Al-Rashid',
        'Jessica Park',
        'Tom Brennan',
        'Yuki Tanaka',
        'Grace Adeyemi',
        'Liam Fletcher',
        'Amara Diallo',
    ]

    name_font = Font(name='Calibri', size=11)
    for row, name in enumerate(staff_names, 2):
        cell = ws.cell(row=row, column=1, value=name)
        cell.font = name_font
        cell.alignment = Alignment(horizontal='left', vertical='center')

    # --- B2:H21: Empty shift assignment grid ---
    # Grid borders to make it visually clear
    thin = Side(style='thin', color='FFB0B0B0')
    grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    grid_align = Alignment(horizontal='center', vertical='center')

    for row in range(2, 22):
        for col in range(2, 9):  # B to H
            cell = ws.cell(row=row, column=col)
            cell.border = grid_border
            cell.alignment = grid_align
            # Leave value empty (no data)

    # --- I2:I21: Empty Total Shifts ---
    for row in range(2, 22):
        cell = ws.cell(row=row, column=9)  # Column I
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- J2:J21: Empty Overtime Flag ---
    for row in range(2, 22):
        cell = ws.cell(row=row, column=10)  # Column J
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Row 22: Empty (will hold COUNTIF totals after task) ---
    ws.cell(row=22, column=1, value='Daily Totals')
    ws.cell(row=22, column=1).font = Font(bold=True, name='Calibri', size=11)

    # --- Column widths ---
    ws.column_dimensions['A'].width = 22   # Staff names
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col_letter].width = 8
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 15

    # --- Row 1 height ---
    ws.row_dimensions[1].height = 20

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet: ShiftSchedule')
    print(f'Staff rows: 2-21 (20 staff members)')
    print(f'Grid: B2:H21 (empty)')
    print(f'Headers: I1=Total Shifts, J1=Overtime Flag')
    print(f'NO data validation, NO formulas, NO conditional formatting')


create_initial()
