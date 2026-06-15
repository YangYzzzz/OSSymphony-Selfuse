"""
Reward Script: VLOOKUP + Pivot Table — Retail Inventory
Task ID: osworld_calc_vlookup_pivot_combined_013
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.50): VLOOKUP formulas fill the Aisle column in Inventory sheet (C2:C11)
  - Component 2 (0.30): Summary sheet has a pivot-style table with Aisle and Total Stock Quantity columns
  - Component 3 (0.20): Summary totals are correct per-aisle values

Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_vlookup_pivot_combined_013'

# Expected summary data: aisle -> total stock quantity
EXPECTED_SUMMARY = {
    'Bakery': 85,
    'Beverages': 138,
    'Canned Goods': 89,
    'Dairy': 456,
    'Grains': 203,
    'Meat': 47,
}

EXPECTED_AISLE_COUNT = 10  # 10 data rows in Inventory (rows 2..11)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # -------------------------------------------------------------------
    # Component 1: VLOOKUP formulas in Inventory Aisle column (0.5 points)
    # The Aisle column (C) in the Inventory sheet should have VLOOKUP
    # formulas for all 10 data rows (C2:C11) that look up from the F:G
    # reference table.
    # This FAILS on initial (all None) and PASSES on golden (all VLOOKUPs).
    # -------------------------------------------------------------------
    try:
        if 'Inventory' not in wb.sheetnames:
            print("FAIL: Component 1 — 'Inventory' sheet not found")
        else:
            ws_inv = wb['Inventory']
            vlookup_count = 0
            for row_idx in range(2, 12):  # rows 2..11 (10 products)
                cell = ws_inv.cell(row=row_idx, column=3)  # Column C
                val = cell.value
                if val and isinstance(val, str) and 'VLOOKUP' in val.upper():
                    vlookup_count += 1

            if vlookup_count == EXPECTED_AISLE_COUNT:
                print(f"PASS: Component 1 — All {vlookup_count}/10 Aisle cells have VLOOKUP formulas (0.5 pts)")
                total_score += 0.5
            elif vlookup_count > 0:
                partial = round(0.5 * vlookup_count / EXPECTED_AISLE_COUNT, 4)
                print(f"PARTIAL: Component 1 — {vlookup_count}/10 Aisle cells have VLOOKUP formulas (+{partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 1 — No VLOOKUP formulas found in Aisle column (C2:C11); found: {vlookup_count}/10")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------
    # Component 2: Summary sheet has pivot-like table structure (0.3 points)
    # The Summary sheet (Sheet2) should have:
    #   - Row 1: headers ("Aisle" and a stock quantity header)
    #   - At least 6 data rows with aisle names
    # This FAILS on initial (Summary sheet is empty) and PASSES on golden.
    # -------------------------------------------------------------------
    try:
        # The task uses 'Summary' as the sheet name (maps to Sheet2)
        summary_sheet_name = None
        for name in wb.sheetnames:
            if name.lower() in ('summary', 'sheet2'):
                summary_sheet_name = name
                break

        if summary_sheet_name is None:
            print("FAIL: Component 2 — No 'Summary' or 'Sheet2' found in workbook")
        else:
            ws_sum = wb[summary_sheet_name]
            # Check that header row exists with aisle-related text
            header_a = ws_sum.cell(row=1, column=1).value
            header_b = ws_sum.cell(row=1, column=2).value

            has_aisle_header = (
                header_a is not None and
                'aisle' in str(header_a).lower()
            )
            has_quantity_header = (
                header_b is not None and
                len(str(header_b).strip()) > 0
            )

            # Count non-empty data rows below header
            data_row_count = 0
            for row_idx in range(2, ws_sum.max_row + 1):
                val_a = ws_sum.cell(row=row_idx, column=1).value
                val_b = ws_sum.cell(row=row_idx, column=2).value
                if val_a is not None and val_b is not None:
                    data_row_count += 1

            if has_aisle_header and has_quantity_header and data_row_count >= 6:
                print(
                    f"PASS: Component 2 — Summary sheet has pivot table: "
                    f"headers ('{header_a}', '{header_b}'), {data_row_count} data rows (0.3 pts)"
                )
                total_score += 0.3
            elif data_row_count >= 6:
                print(
                    f"PARTIAL: Component 2 — Summary sheet has {data_row_count} data rows "
                    f"but missing proper headers (header_a={header_a!r}, header_b={header_b!r}) (+0.15 pts)"
                )
                total_score += 0.15
            else:
                print(
                    f"FAIL: Component 2 — Summary sheet has only {data_row_count} data rows "
                    f"(need >=6). Headers: ('{header_a}', '{header_b}')"
                )
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------
    # Component 3: Summary aisle totals are correct (0.2 points)
    # Each aisle's total stock quantity must match the expected values
    # computed from the reference data in the Inventory sheet.
    # Dairy=456, Grains=203, Meat=47, Bakery=85, Beverages=138, Canned Goods=89
    # This FAILS on initial (Summary sheet is empty) and PASSES on golden.
    # -------------------------------------------------------------------
    try:
        summary_sheet_name = None
        for name in wb.sheetnames:
            if name.lower() in ('summary', 'sheet2'):
                summary_sheet_name = name
                break

        if summary_sheet_name is None:
            print("FAIL: Component 3 — No Summary/Sheet2 sheet found")
        else:
            ws_sum = wb[summary_sheet_name]

            # Build a dict of aisle -> stock total from the sheet
            found_totals = {}
            for row_idx in range(2, ws_sum.max_row + 1):
                aisle_val = ws_sum.cell(row=row_idx, column=1).value
                qty_val = ws_sum.cell(row=row_idx, column=2).value
                if aisle_val is not None and qty_val is not None:
                    try:
                        found_totals[str(aisle_val).strip()] = float(qty_val)
                    except (ValueError, TypeError):
                        pass

            # Compare against expected totals
            correct_count = 0
            for aisle, expected_qty in EXPECTED_SUMMARY.items():
                actual_qty = found_totals.get(aisle)
                if actual_qty is not None and abs(actual_qty - expected_qty) <= 0.5:
                    correct_count += 1
                else:
                    print(
                        f"  FAIL: Component 3 — Aisle '{aisle}': "
                        f"expected {expected_qty}, found {actual_qty}"
                    )

            if correct_count == len(EXPECTED_SUMMARY):
                print(f"PASS: Component 3 — All {correct_count}/6 aisle totals are correct (0.2 pts)")
                total_score += 0.2
            elif correct_count > 0:
                partial = round(0.2 * correct_count / len(EXPECTED_SUMMARY), 4)
                print(
                    f"PARTIAL: Component 3 — {correct_count}/6 aisle totals correct (+{partial} pts)"
                )
                total_score += partial
            else:
                print(
                    f"FAIL: Component 3 — No aisle totals matched. Found: {found_totals}"
                )
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {round(final_score, 4)}")
    return final_score


# Entry point — test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
