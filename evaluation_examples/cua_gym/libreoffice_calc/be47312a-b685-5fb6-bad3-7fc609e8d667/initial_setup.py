"""
Initial Setup: Retail inventory dataset with VLOOKUP + pivot table task
Task ID: osworld_calc_vlookup_pivot_combined_013
Domain: libreoffice_calc

Creates Sheet1 with inventory data (Product ID, Product Name, Aisle [empty], Stock Quantity)
and a store layout reference in columns F-G. Sheet2 is blank (pivot table to be added by agent).
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_vlookup_pivot_combined_013'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Inventory ---
    ws1 = wb.active
    ws1.title = 'Inventory'

    # Headers in columns A-D (inventory) and F-G (store layout reference)
    inventory_headers = ['Product ID', 'Product Name', 'Aisle', 'Stock Quantity']
    for col, h in enumerate(inventory_headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    ref_headers = ['Product ID', 'Aisle']
    for col, h in enumerate(ref_headers, 6):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Inventory data: Product ID, Product Name, Aisle (EMPTY - to be filled via VLOOKUP), Stock Quantity
    inventory_data = [
        ['P001', 'Whole Milk 1L',          None, 120],
        ['P002', 'Sourdough Bread',         None, 85],
        ['P003', 'Cheddar Cheese 400g',     None, 64],
        ['P004', 'Free Range Eggs x12',     None, 200],
        ['P005', 'Organic Chicken Breast',  None, 47],
        ['P006', 'Basmati Rice 2kg',        None, 93],
        ['P007', 'Orange Juice 2L',         None, 138],
        ['P008', 'Greek Yogurt 500g',       None, 72],
        ['P009', 'Pasta Fusilli 500g',      None, 110],
        ['P010', 'Tomato Sauce Jar',        None, 89],
    ]
    for r, row_data in enumerate(inventory_data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Store layout reference in columns F-G (maps Product ID to Aisle)
    store_layout = [
        ['P001', 'Dairy'],
        ['P002', 'Bakery'],
        ['P003', 'Dairy'],
        ['P004', 'Dairy'],
        ['P005', 'Meat'],
        ['P006', 'Grains'],
        ['P007', 'Beverages'],
        ['P008', 'Dairy'],
        ['P009', 'Grains'],
        ['P010', 'Canned Goods'],
    ]
    for r, row_data in enumerate(store_layout, 2):
        for c, val in enumerate(row_data, 6):
            ws1.cell(row=r, column=c, value=val)

    # Adjust column widths for readability
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 28
    ws1.column_dimensions['C'].width = 14
    ws1.column_dimensions['D'].width = 16
    ws1.column_dimensions['F'].width = 12
    ws1.column_dimensions['G'].width = 14

    # --- Sheet 2: Summary (blank - pivot table to be created by agent) ---
    ws2 = wb.create_sheet('Summary')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
