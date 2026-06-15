"""
Initial Setup: HR Salary Currency Format Task
Task ID: calc_hr_salary_currency_format_003
Domain: libreoffice_calc

Creates a Compensation sheet with employee data:
- Column E (Annual Salary) and F (Bonus) contain raw numbers with NO formatting
- Column G (Total Comp) is EMPTY
- Sheet is UNPROTECTED with all cells UNLOCKED
"""

import openpyxl
from openpyxl.styles import Protection

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_salary_currency_format_003'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Compensation'

    # Row 1: Headers
    headers = ['Emp ID', 'Name', 'Department', 'Title', 'Annual Salary', 'Bonus', 'Total Comp']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic employee data: 66 rows (rows 2-67)
    employees = [
        ('E001', 'Sarah Chen', 'Engineering', 'Senior Software Engineer', 110000, 10000),
        ('E002', 'Marcus Johnson', 'Marketing', 'Marketing Manager', 92500, 8000),
        ('E003', 'Priya Patel', 'Finance', 'Financial Analyst', 78000, 6000),
        ('E004', 'James O\'Brien', 'HR', 'HR Business Partner', 72000, 5500),
        ('E005', 'Mei Lin', 'Engineering', 'Software Engineer II', 95000, 9000),
        ('E006', 'David Rodriguez', 'Sales', 'Sales Representative', 65000, 12000),
        ('E007', 'Amanda Foster', 'Legal', 'Corporate Counsel', 125000, 11000),
        ('E008', 'Kevin Nakamura', 'Engineering', 'DevOps Engineer', 105000, 9500),
        ('E009', 'Lisa Thompson', 'Marketing', 'Brand Strategist', 83000, 7000),
        ('E010', 'Robert Kim', 'Finance', 'Senior Financial Analyst', 91000, 8500),
        ('E011', 'Angela White', 'Operations', 'Operations Manager', 88000, 8000),
        ('E012', 'Thomas Green', 'Engineering', 'Principal Engineer', 135000, 13000),
        ('E013', 'Rachel Adams', 'HR', 'Recruiter', 62000, 4500),
        ('E014', 'Carlos Mendez', 'Sales', 'Account Executive', 71000, 15000),
        ('E015', 'Stephanie Lee', 'Engineering', 'QA Engineer', 82000, 7500),
        ('E016', 'Brian Turner', 'Finance', 'Controller', 115000, 10500),
        ('E017', 'Michelle Clark', 'Marketing', 'Content Manager', 75000, 6000),
        ('E018', 'Jason Park', 'Engineering', 'Machine Learning Engineer', 128000, 12000),
        ('E019', 'Nicole Harris', 'Legal', 'Paralegal', 58000, 4000),
        ('E020', 'Derek Brown', 'Sales', 'Regional Sales Manager', 98000, 20000),
        ('E021', 'Vanessa Moore', 'Operations', 'Supply Chain Analyst', 74000, 6500),
        ('E022', 'Christopher Scott', 'Engineering', 'Backend Engineer', 102000, 9200),
        ('E023', 'Tiffany Hall', 'HR', 'Benefits Coordinator', 60000, 4200),
        ('E024', 'Anthony Rivera', 'Finance', 'Treasury Analyst', 81000, 7200),
        ('E025', 'Emma Wilson', 'Marketing', 'Digital Marketing Specialist', 70000, 5800),
        ('E026', 'Nathan Hughes', 'Engineering', 'Frontend Engineer', 99000, 9000),
        ('E027', 'Olivia Baker', 'Sales', 'Inside Sales Rep', 57000, 11000),
        ('E028', 'Joshua Mitchell', 'Operations', 'Logistics Coordinator', 63000, 5000),
        ('E029', 'Samantha Price', 'Engineering', 'Data Engineer', 112000, 10200),
        ('E030', 'Dylan Carter', 'Legal', 'Contract Specialist', 76000, 6200),
        ('E031', 'Grace Nguyen', 'Finance', 'Accounts Payable Specialist', 55000, 3800),
        ('E032', 'Ian Campbell', 'Engineering', 'Cloud Architect', 142000, 14000),
        ('E033', 'Diana Evans', 'Marketing', 'SEO Manager', 78500, 6500),
        ('E034', 'Marcus Bell', 'Sales', 'Sales Engineer', 105000, 18000),
        ('E035', 'Alexis Torres', 'HR', 'HR Generalist', 66000, 5200),
        ('E036', 'Patrick Peterson', 'Operations', 'Process Improvement Lead', 87000, 7800),
        ('E037', 'Chloe Sanchez', 'Engineering', 'Security Engineer', 118000, 11000),
        ('E038', 'Brandon Morris', 'Finance', 'Tax Analyst', 79000, 6800),
        ('E039', 'Natalie Rogers', 'Marketing', 'PR Specialist', 68000, 5500),
        ('E040', 'Victor Reed', 'Sales', 'Business Development Rep', 73000, 13500),
        ('E041', 'Megan Cook', 'Operations', 'Project Manager', 93000, 8200),
        ('E042', 'Tyler Morgan', 'Engineering', 'Embedded Systems Engineer', 107000, 9800),
        ('E043', 'Brittany Bailey', 'Legal', 'Compliance Officer', 89000, 7900),
        ('E044', 'Nicholas Rivera', 'Finance', 'Budget Analyst', 77000, 6700),
        ('E045', 'Kayla Cooper', 'HR', 'Talent Acquisition Specialist', 68500, 5300),
        ('E046', 'Ryan Richardson', 'Engineering', 'Platform Engineer', 114000, 10500),
        ('E047', 'Ashley Cox', 'Marketing', 'Social Media Manager', 65000, 5100),
        ('E048', 'Jordan Howard', 'Sales', 'VP of Sales', 165000, 35000),
        ('E049', 'Taylor Ward', 'Operations', 'Quality Assurance Manager', 86000, 7600),
        ('E050', 'Morgan Torres', 'Engineering', 'Technical Lead', 130000, 12500),
        ('E051', 'Casey Peterson', 'Finance', 'CFO', 185000, 25000),
        ('E052', 'Riley Gray', 'Marketing', 'VP of Marketing', 145000, 15000),
        ('E053', 'Avery James', 'HR', 'VP of HR', 138000, 13500),
        ('E054', 'Drew Watson', 'Legal', 'General Counsel', 175000, 18000),
        ('E055', 'Cameron Brooks', 'Engineering', 'VP of Engineering', 178000, 20000),
        ('E056', 'Hayden Kelly', 'Sales', 'Sales Director', 122000, 22000),
        ('E057', 'Blake Sanders', 'Operations', 'COO', 195000, 28000),
        ('E058', 'Quinn Price', 'Engineering', 'Software Engineer I', 82000, 6500),
        ('E059', 'Peyton Bennett', 'Marketing', 'Marketing Analyst', 64000, 5000),
        ('E060', 'Reese Wood', 'Finance', 'Payroll Specialist', 59000, 4200),
        ('E061', 'Sidney Barnes', 'HR', 'HRIS Analyst', 72500, 5800),
        ('E062', 'Dakota Ross', 'Engineering', 'Full Stack Developer', 104000, 9400),
        ('E063', 'Sage Henderson', 'Sales', 'Account Manager', 78000, 14000),
        ('E064', 'River Coleman', 'Operations', 'Facilities Manager', 69000, 5600),
        ('E065', 'Shiloh Jenkins', 'Legal', 'IP Attorney', 148000, 14500),
        ('E066', 'Phoenix Perry', 'Finance', 'Investment Analyst', 96000, 8800),
    ]

    for r, emp in enumerate(employees, 2):
        emp_id, name, dept, title, salary, bonus = emp
        ws.cell(row=r, column=1, value=emp_id)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=dept)
        ws.cell(row=r, column=4, value=title)
        ws.cell(row=r, column=5, value=salary)   # raw number, no formatting
        ws.cell(row=r, column=6, value=bonus)    # raw number, no formatting
        # Column G (Total Comp) intentionally left empty

    # Ensure all cells are unlocked (sheet not protected)
    # By default openpyxl creates unlocked cells — we confirm by not setting any Protection
    # The sheet is NOT protected

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Rows: 1 header + 66 data rows = 67 total')
    print(f'Column E: raw salary numbers (no currency format)')
    print(f'Column F: raw bonus numbers (no currency format)')
    print(f'Column G: empty (no formulas)')
    print(f'Sheet protection: NONE')


create_initial()
