"""
Reward Script: Format salary columns as USD currency, add Total Comp formulas, protect salary columns.
Task ID: calc_hr_salary_currency_format_003
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): E2:E67 and F2:F67 formatted as $#,##0.00 currency
  Component 2 (0.30): G2:G67 contain =E{n}+F{n} formulas (row-adjusted)
  Component 3 (0.10): G2:G67 formatted as $#,##0.00 currency
  Component 4 (0.15): E1:F67 cells are locked, A-D and G cells are unlocked
  Component 5 (0.15): Sheet protection is enabled
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_salary_currency_format_003'
EXPECTED_FORMAT = '$#,##0.00'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the sheet exists
    if 'Compensation' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Compensation' not found in workbook.")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Compensation']

    # Component 1: E2:E67 and F2:F67 are formatted as $#,##0.00 currency (0.30 points)
    # Task: "format them as USD currency with comma separators and two decimal places"
    # Initial: All cells have 'General' format
    # Golden: E2:E67 and F2:F67 have '$#,##0.00' format
    try:
        e_formatted = 0
        f_formatted = 0
        e_total = 66  # rows 2-67
        f_total = 66

        for row in range(2, 68):
            e_fmt = ws.cell(row=row, column=5).number_format
            f_fmt = ws.cell(row=row, column=6).number_format
            if e_fmt == EXPECTED_FORMAT:
                e_formatted += 1
            if f_fmt == EXPECTED_FORMAT:
                f_formatted += 1

        if e_formatted == e_total and f_formatted == f_total:
            print(f"PASS: Component 1 — E2:E67 and F2:F67 all formatted as {EXPECTED_FORMAT!r} (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 — E formatted: {e_formatted}/{e_total}, F formatted: {f_formatted}/{f_total}; expected {EXPECTED_FORMAT!r}")
            # Partial credit within this component
            if e_formatted == e_total:
                print(f"  PARTIAL: E column fully formatted, F not ({f_formatted}/{f_total}) — no partial pts for split")
            elif f_formatted == f_total:
                print(f"  PARTIAL: F column fully formatted, E not ({e_formatted}/{e_total}) — no partial pts for split")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: G2:G67 contain row-adjusted =E{n}+F{n} formulas (0.30 points)
    # Task: "Cells G2:G67 must contain formula =E2+F2 (adjusted per row)"
    # Initial: G column is empty (all None)
    # Golden: Each G cell contains =E{row}+F{row}
    try:
        g_formula_correct = 0
        g_total = 66  # rows 2-67
        first_fail = None

        for row in range(2, 68):
            val = ws.cell(row=row, column=7).value
            expected_formula = f'=E{row}+F{row}'
            # Accept both upper and lower case
            if isinstance(val, str) and val.strip().upper() == expected_formula.upper():
                g_formula_correct += 1
            else:
                if first_fail is None:
                    first_fail = (row, val)

        if g_formula_correct == g_total:
            print(f"PASS: Component 2 — G2:G67 all have row-adjusted =En+Fn formulas (0.30 pts)")
            total_score += 0.30
        else:
            pct = g_formula_correct / g_total
            if first_fail:
                print(f"FAIL: Component 2 — {g_formula_correct}/{g_total} rows correct; first fail at G{first_fail[0]}: {first_fail[1]!r}")
            else:
                print(f"FAIL: Component 2 — 0/{g_total} G cells have correct formula")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: G2:G67 are formatted as $#,##0.00 currency (0.10 points)
    # Task: "Cells G2:G67 must contain formula =E2+F2 (adjusted per row) and also formatted as $#,##0.00"
    # Initial: G column cells have 'General' format
    # Golden: G2:G67 have '$#,##0.00' format
    try:
        g_formatted = 0
        g_fmt_total = 66

        for row in range(2, 68):
            g_fmt = ws.cell(row=row, column=7).number_format
            if g_fmt == EXPECTED_FORMAT:
                g_formatted += 1

        if g_formatted == g_fmt_total:
            print(f"PASS: Component 3 — G2:G67 all formatted as {EXPECTED_FORMAT!r} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — G formatted: {g_formatted}/{g_fmt_total} with {EXPECTED_FORMAT!r}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: E1:F67 cells are locked, A-D and G columns are unlocked (0.15 points)
    # Task: "Columns E and F (E1:F67) must be locked ... Columns A-D and G must remain unprotected and editable"
    # Initial: All cells default locked=True but sheet protection is off (locking meaningless)
    # Golden: E and F cells locked=True, A-D and G cells locked=False, sheet protection on
    try:
        ef_locked_correct = 0
        ef_locked_total = 0
        adg_unlocked_correct = 0
        adg_unlocked_total = 0

        # Check E1:F67 are locked
        for row in range(1, 68):
            for col in (5, 6):  # E=5, F=6
                cell = ws.cell(row=row, column=col)
                ef_locked_total += 1
                if cell.protection.locked:
                    ef_locked_correct += 1

        # Check A-D (cols 1-4) and G (col 7) rows 2-67 are unlocked
        for row in range(2, 68):
            for col in (1, 2, 3, 4, 7):  # A, B, C, D, G
                cell = ws.cell(row=row, column=col)
                adg_unlocked_total += 1
                if not cell.protection.locked:
                    adg_unlocked_correct += 1

        ef_ok = ef_locked_correct == ef_locked_total
        adg_ok = adg_unlocked_correct == adg_unlocked_total

        if ef_ok and adg_ok:
            print(f"PASS: Component 4 — E1:F67 all locked ({ef_locked_correct}/{ef_locked_total}), A-D and G unlocked ({adg_unlocked_correct}/{adg_unlocked_total}) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — E1:F67 locked: {ef_locked_correct}/{ef_locked_total}, A-D/G unlocked: {adg_unlocked_correct}/{adg_unlocked_total}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Sheet protection is enabled (0.15 points)
    # Task: "protect the salary column so nobody can accidentally edit it" with password 'HROnly2024'
    # Initial: ws.protection.sheet = False (unprotected)
    # Golden: ws.protection.sheet = True (protected with password)
    try:
        sheet_protected = ws.protection.sheet
        has_password = ws.protection.password is not None and ws.protection.password != ''

        if sheet_protected:
            print(f"PASS: Component 5 — Sheet protection is enabled (password hash: {ws.protection.password!r}) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 5 — Sheet protection is NOT enabled (ws.protection.sheet={sheet_protected})")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
