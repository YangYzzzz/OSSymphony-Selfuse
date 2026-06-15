"""
Initial Setup: Add data validation to B2:B30 for whole numbers 2020-2030
Task ID: calc_dop_validate_integer_062
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_dop_validate_integer_062'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: FiscalReports ---
    ws = wb.active
    ws.title = 'FiscalReports'

    # Headers
    headers = ['Report ID', 'Fiscal Year', 'Quarter', 'Division', 'Revenue']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Data rows 2-30 (29 rows)
    # Include some invalid values: 2019, 2031, and 2-digit years like 25
    divisions = ['Operations', 'Marketing', 'Finance', 'Engineering', 'Sales',
                 'HR', 'Legal', 'IT', 'Procurement', 'Research']
    quarters = ['Q1', 'Q2', 'Q3', 'Q4']

    data = [
        ['RPT-001', 2023, 'Q1', 'Operations',    1542300],
        ['RPT-002', 2019, 'Q2', 'Marketing',      893450],  # invalid: 2019
        ['RPT-003', 2024, 'Q3', 'Finance',        2310780],
        ['RPT-004', 2022, 'Q4', 'Engineering',    1875600],
        ['RPT-005', 2031, 'Q1', 'Sales',           765400],  # invalid: 2031
        ['RPT-006', 2025, 'Q2', 'HR',              432100],
        ['RPT-007', 2021, 'Q3', 'Legal',           578900],
        ['RPT-008', 25,   'Q4', 'IT',              991200],  # invalid: 2-digit
        ['RPT-009', 2026, 'Q1', 'Procurement',    1234500],
        ['RPT-010', 2020, 'Q2', 'Research',       1678900],
        ['RPT-011', 2023, 'Q3', 'Operations',     2045600],
        ['RPT-012', 2024, 'Q4', 'Marketing',       934700],
        ['RPT-013', 2022, 'Q1', 'Finance',        1567800],
        ['RPT-014', 2021, 'Q2', 'Engineering',     889300],
        ['RPT-015', 2028, 'Q3', 'Sales',          1120400],
        ['RPT-016', 2027, 'Q4', 'HR',              654200],
        ['RPT-017', 2030, 'Q1', 'Legal',           743800],
        ['RPT-018', 2019, 'Q2', 'IT',              512300],  # invalid: 2019
        ['RPT-019', 2025, 'Q3', 'Procurement',    1890600],
        ['RPT-020', 2026, 'Q4', 'Research',       2234100],
        ['RPT-021', 2023, 'Q1', 'Operations',     1456700],
        ['RPT-022', 2024, 'Q2', 'Marketing',      1098500],
        ['RPT-023', 25,   'Q3', 'Finance',         678400],  # invalid: 2-digit
        ['RPT-024', 2022, 'Q4', 'Engineering',    1345900],
        ['RPT-025', 2029, 'Q1', 'Sales',           923700],
        ['RPT-026', 2020, 'Q2', 'HR',              567800],
        ['RPT-027', 2021, 'Q3', 'Legal',           812300],
        ['RPT-028', 2031, 'Q4', 'IT',             1567200],  # invalid: 2031
        ['RPT-029', 2027, 'Q1', 'Procurement',    1789400],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # No data validation on column B (this is the initial state)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
