"""
Reward Script: Import CSV into LibreOffice Calc with comma delimiter
Task ID: calc_gsi_013
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Correct header row (Product Code, Description, Quantity, Unit Price)
  Component 2 (0.25): Correct number of data rows (15 rows)
  Component 3 (0.25): Numeric types for Quantity (int) and Unit Price (float) columns
  Component 4 (0.25): Spot-check data values match the CSV source
"""

import os

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_013'


def verify_task(file_path):
    import openpyxl
    """
    Verify that the CSV was correctly imported into an xlsx file.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Use the first sheet regardless of name
    ws = wb.worksheets[0]

    # Component 1: Correct header row (0.25 points)
    try:
        expected_headers = ['Product Code', 'Description', 'Quantity', 'Unit Price']
        actual_headers = []
        for col in range(1, 5):
            val = ws.cell(row=1, column=col).value
            actual_headers.append(str(val).strip() if val is not None else None)

        if actual_headers == expected_headers:
            print(f"PASS: Component 1 — Headers match: {actual_headers} (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — Expected headers {expected_headers}, found {actual_headers}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Correct number of data rows - 15 rows (0.25 points)
    try:
        # Count non-empty rows starting from row 2
        data_row_count = 0
        for row_idx in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=1).value
            if cell_val is not None and str(cell_val).strip() != '':
                data_row_count += 1

        if data_row_count == 15:
            print(f"PASS: Component 2 — Found 15 data rows as expected (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — Expected 15 data rows, found {data_row_count}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Numeric types for Quantity and Unit Price columns (0.25 points)
    # Quantity (col C) should be int/float, Unit Price (col D) should be float
    try:
        numeric_issues = []

        # Check a sample of rows for numeric types
        check_rows = [2, 5, 9, 13, 16]  # spread across data
        for r in check_rows:
            qty_val = ws.cell(row=r, column=3).value
            price_val = ws.cell(row=r, column=4).value

            if qty_val is not None and not isinstance(qty_val, (int, float)):
                numeric_issues.append(f"Row {r} Quantity is {type(qty_val).__name__}: {qty_val}")

            if price_val is not None and not isinstance(price_val, (int, float)):
                numeric_issues.append(f"Row {r} Unit Price is {type(price_val).__name__}: {price_val}")

        if len(numeric_issues) == 0:
            print(f"PASS: Component 3 — Quantity and Unit Price columns are numeric types (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 3 — Non-numeric values found: {numeric_issues}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Spot-check data values match CSV source (0.25 points)
    try:
        # Expected values from the CSV (row, col, expected_value)
        spot_checks = [
            (2, 1, 'WH-1001'),       # First product code
            (2, 2, 'Industrial Bolt M10x50'),  # First description
            (2, 3, 2450),             # First quantity
            (2, 4, 0.85),             # First unit price
            (9, 1, 'WH-1008'),       # Mid product code
            (9, 3, 185),              # Mid quantity
            (9, 4, 8.75),             # Mid unit price (highest price)
            (16, 1, 'WH-1015'),      # Last product code
            (16, 2, 'Nylon Lock Nut M8'),  # Last description
            (16, 3, 3400),            # Last quantity
            (16, 4, 0.38),            # Last unit price
        ]

        checks_passed = 0
        for row, col, expected in spot_checks:
            actual = ws.cell(row=row, column=col).value
            if isinstance(expected, (int, float)) and isinstance(actual, (int, float)):
                if abs(float(actual) - float(expected)) < 0.01:
                    checks_passed += 1
                else:
                    print(f"  DETAIL: Row {row}, Col {col}: expected {expected}, got {actual}")
            elif actual is not None and str(actual).strip() == str(expected).strip():
                checks_passed += 1
            else:
                print(f"  DETAIL: Row {row}, Col {col}: expected {expected!r}, got {actual!r}")

        # Require at least 9 out of 11 checks to pass for full credit
        if checks_passed >= 9:
            print(f"PASS: Component 4 — {checks_passed}/11 spot checks passed (0.25 pts)")
            total_score += 0.25
        elif checks_passed >= 6:
            partial = round(0.25 * (checks_passed / 11), 2)
            print(f"PARTIAL: Component 4 — {checks_passed}/11 spot checks passed ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {checks_passed}/11 spot checks passed")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
