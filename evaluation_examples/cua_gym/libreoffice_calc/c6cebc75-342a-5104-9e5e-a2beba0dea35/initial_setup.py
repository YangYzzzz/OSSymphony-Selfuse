"""
Initial Setup: Numbers stored as text in column A with SUM formula showing 0
Task ID: calc_tbl_016
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_016'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Header
    ws.cell(row=1, column=1, value="Values")
    ws.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)

    # A2:A100 — 99 numbers stored as TEXT strings
    # Generate a realistic mix of integers and decimals
    values = []
    for i in range(99):
        if random.random() < 0.7:
            # Integer-like values
            val = random.randint(10, 500)
            values.append(str(val))
        else:
            # Decimal values
            val = round(random.uniform(10, 500), 2)
            values.append(str(val))

    for idx, text_val in enumerate(values):
        cell = ws.cell(row=idx + 2, column=1)
        # Store as explicit text string — this is the key:
        # setting value to a string that looks like a number
        cell.value = text_val
        # Force the cell to be text format so LibreOffice shows green triangles
        cell.number_format = '@'

    # A101: SUM formula — will display 0 because values are stored as text
    ws.cell(row=101, column=1, value="=SUM(A2:A100)")

    # Set column width for readability
    ws.column_dimensions['A'].width = 15

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
