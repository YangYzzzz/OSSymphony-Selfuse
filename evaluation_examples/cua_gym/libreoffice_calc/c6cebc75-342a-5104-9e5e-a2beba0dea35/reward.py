"""
Reward Script: Convert text-stored numbers to actual numbers in column A
Task ID: calc_tbl_016
Domain: libreoffice_calc
Scoring:
  Component 1 (0.6): Proportion of A2:A100 cells that are numeric type (int/float)
  Component 2 (0.4): Sum of numeric values in A2:A100 matches expected total (~23590.96)
"""

import os
import time


WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_016'
EXPECTED_SUM = 23590.96
SUM_TOLERANCE = 1.0  # allow small floating point differences


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Component 1: Proportion of A2:A100 that are numeric (0.6 points)
    # Initial state: all 99 cells are strings -> 0 numeric -> 0.0 pts
    # Golden state: all 99 cells are numeric -> 99/99 -> 0.6 pts
    try:
        numeric_count = 0
        total_cells = 99  # A2:A100
        for r in range(2, 101):
            val = ws.cell(row=r, column=1).value
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                numeric_count += 1
        proportion = numeric_count / total_cells
        component1 = round(proportion * 0.6, 4)
        if numeric_count == total_cells:
            print(f"PASS: Component 1 — All {numeric_count}/{total_cells} cells are numeric ({component1} pts)")
            total_score += component1
        elif numeric_count > 0:
            print(f"PARTIAL: Component 1 — {numeric_count}/{total_cells} cells are numeric ({component1} pts)")
            total_score += component1
        else:
            print(f"FAIL: Component 1 — 0/{total_cells} cells are numeric (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Sum of numeric values matches expected total (0.4 points)
    # Initial state: all strings, numeric sum = 0 -> FAIL
    # Golden state: all numeric, sum ~23590.96 -> PASS
    try:
        numeric_sum = 0.0
        for r in range(2, 101):
            val = ws.cell(row=r, column=1).value
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                numeric_sum += float(val)
        if abs(numeric_sum - EXPECTED_SUM) <= SUM_TOLERANCE:
            print(f"PASS: Component 2 — Numeric sum {numeric_sum:.2f} matches expected {EXPECTED_SUM:.2f} (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 2 — Numeric sum {numeric_sum:.2f} != expected {EXPECTED_SUM:.2f}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
