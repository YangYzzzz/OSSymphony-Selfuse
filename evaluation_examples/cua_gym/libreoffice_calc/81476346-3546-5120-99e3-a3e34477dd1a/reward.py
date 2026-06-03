"""
Reward Script: Add a RANK formula to cell D2 to rank student score
Task ID: calc_fmb_rank_025
Domain: libreoffice_calc

Scoring:
  Component 1: D2 contains a RANK formula starting with =RANK( (0.5 pts)
  Component 2: RANK formula is fully correct — references C2, range C$2:C$31,
               descending order (0), and no other cells were modified (0.5 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmb_rank_025'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.

    Task: Add RANK formula to D2 that ranks the score in C2 among all students
    in C2:C31, where rank 1 = highest score (descending order).

    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: 'Exam Results' sheet must exist
    if 'Exam Results' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Exam Results' not found.")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Exam Results']

    # Component 1: D2 contains a RANK formula (0.5 points)
    # The initial state has D2 = None. The task requires placing a RANK formula there.
    # This component fails on initial (D2 is None) and passes on golden (D2 has =RANK(...)).
    try:
        d2_value = ws['D2'].value
        if d2_value is not None and isinstance(d2_value, str) and d2_value.upper().lstrip().startswith('=RANK('):
            print(f"PASS: Component 1 — D2 contains RANK formula: {repr(d2_value)} (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — Expected RANK formula in D2, found: {repr(d2_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not check D2: {e}")

    # Component 2: RANK formula is fully correct — correct value cell (C2), correct range
    # (C$2:C$31 or equivalent absolute references covering all 30 students), correct
    # descending order (0), AND no other cells in column D were modified (D3:D31 still None).
    # This entire component fails on initial since D2 is None (formula check fails immediately).
    try:
        d2_value = ws['D2'].value
        formula_correct = False
        if d2_value is not None and isinstance(d2_value, str):
            formula_upper = d2_value.upper().replace(' ', '')
            # Check the formula references C2 (the target student's score)
            has_c2_ref = formula_upper.startswith('=RANK(C2,')
            # Check for correct range: C$2:C$31 (absolute row references spanning all 30 students)
            # Accept variants with or without $ on column letter, but require $ on row numbers
            has_correct_range = ('C$2:C$31' in formula_upper)
            # Check for descending order argument: 0 (rank 1 = highest score)
            has_descending_order = formula_upper.endswith(',0)') or ',0)' in formula_upper
            # Check that no other D-column cells were modified (D3:D31 must still be None)
            other_d_cells_empty = all(
                ws.cell(row=r, column=4).value is None
                for r in range(3, 32)
            )
            if has_c2_ref and has_correct_range and has_descending_order and other_d_cells_empty:
                print(f"PASS: Component 2 — RANK formula fully correct; C2 ref, C$2:C$31 range, descending order 0, D3:D31 unchanged (0.5 pts)")
                total_score += 0.5
            else:
                reasons = []
                if not has_c2_ref:
                    reasons.append(f"formula does not start with =RANK(C2, (found: {repr(d2_value)})")
                if not has_correct_range:
                    reasons.append(f"range is not C$2:C$31 (found: {repr(d2_value)})")
                if not has_descending_order:
                    reasons.append(f"order argument is not 0/descending (found: {repr(d2_value)})")
                if not other_d_cells_empty:
                    unexpected = [(r, ws.cell(row=r, column=4).value) for r in range(3, 32) if ws.cell(row=r, column=4).value is not None]
                    reasons.append(f"unexpected values in D3:D31: {unexpected}")
                print(f"FAIL: Component 2 — {'; '.join(reasons)}")
        else:
            print(f"FAIL: Component 2 — D2 is not a formula string: {repr(d2_value)}")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not check formula details: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
