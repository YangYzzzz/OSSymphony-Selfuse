"""
Initial Setup: Add a rank column to the student scores table
Task ID: calc_fmb_rank_025
Domain: libreoffice_calc

Creates a spreadsheet with one sheet 'Exam Results' containing:
- 30 students with Student ID, Name, Score columns
- An empty D column (Rank) with header
- C2=88, other scores as specified in context
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_rank_025'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Exam Results ---
    ws = wb.active
    ws.title = 'Exam Results'

    # Row 1: Headers
    ws['A1'] = 'Student ID'
    ws['B1'] = 'Name'
    ws['C1'] = 'Score'
    ws['D1'] = 'Rank'

    # 30 students data
    # C2=88, remaining scores (29 values): 92, 74, 68, 95, 83, 77, 91, 65, 72, 84,
    # 79, 96, 61, 88, 75, 93, 82, 70, 86, 78, 94, 67, 89, 81, 76, 90, 73, 85, 71
    scores = [
        88,  # C2 — student being ranked (row 2)
        92, 74, 68, 95, 83, 77, 91, 65, 72,
        84, 79, 96, 61, 88, 75, 93, 82, 70,
        86, 78, 94, 67, 89, 81, 76, 90, 73,
        85, 71
    ]

    # Realistic student names
    names = [
        'Emma Hartwell',
        'Liam Nakamura',
        'Sofia Delgado',
        'Noah Okonkwo',
        'Ava Kowalski',
        'James Patel',
        'Mia Lindqvist',
        'Oliver Ferretti',
        'Chloe Mbeki',
        'Ethan Sorensen',
        'Isabella Yamamoto',
        'Lucas Hoffman',
        'Amelia Tremblay',
        'Benjamin Osei',
        'Harper Vasquez',
        'Alexander Petrov',
        'Evelyn Nguyen',
        'William Abramson',
        'Abigail Castillo',
        'Henry Johansson',
        'Emily Rashid',
        'Michael Andersson',
        'Elizabeth Okello',
        'Daniel Bergstrom',
        'Sofia Mensah',
        'Matthew Eriksen',
        'Avery Tanaka',
        'Jackson Ferreira',
        'Scarlett Holmberg',
        'Sebastian Adeyemi',
    ]

    for i, (name, score) in enumerate(zip(names, scores), start=2):
        student_id = f'STU{1000 + i - 2:04d}'
        ws.cell(row=i, column=1, value=student_id)
        ws.cell(row=i, column=2, value=name)
        ws.cell(row=i, column=3, value=score)
        # Column D (Rank) intentionally left empty

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet: Exam Results')
    print(f'Rows: 31 (1 header + 30 data)')
    print(f'D2 is empty (target cell for rank formula)')


create_initial()
