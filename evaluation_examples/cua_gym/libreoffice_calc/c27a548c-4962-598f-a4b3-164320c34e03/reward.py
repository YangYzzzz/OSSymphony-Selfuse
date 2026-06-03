"""
Reward Script: Multi-drop delivery route log with on-time compliance tracking
Task ID: calc_ops_fleet_route_log_070
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.25): RouteStops I2:I201 — IF(AND(H>=F, H<=G), "On Time", "Late") formulas
  - Component 2 (0.20): RouteStops J2:J201 — IF(I="Late", MAX(0,(H-G)*1440), 0) formulas
  - Component 3 (0.15): DriverPerformance B2:B9 — COUNTIF(RouteStops!B, driver) formulas
  - Component 4 (0.15): DriverPerformance C2:C9 — COUNTIFS(B, driver, I, "On Time") formulas
  - Component 5 (0.10): DriverPerformance D2:D9 — =C/B percentage formulas with 0.00% format
  - Component 6 (0.10): DriverPerformance E2:E9 — AVERAGEIF(RouteStops!B, driver, J) formulas
  - Component 7 (0.05): Conditional formatting on D2:D9 — red <90%, green >=95%
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — reward script runs on VM
TASK_ID = 'calc_ops_fleet_route_log_070'


def normalize_formula(f):
    """Normalize formula: strip leading =, uppercase, remove spaces."""
    if not f:
        return ''
    f = str(f).strip()
    if f.startswith('='):
        f = f[1:]
    return f.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: required sheets exist
    if 'RouteStops' not in wb.sheetnames:
        print("CRITICAL: Sheet 'RouteStops' not found")
        print("REWARD: 0.0")
        return 0.0
    if 'DriverPerformance' not in wb.sheetnames:
        print("CRITICAL: Sheet 'DriverPerformance' not found")
        print("REWARD: 0.0")
        return 0.0

    ws_rs = wb['RouteStops']
    ws_dp = wb['DriverPerformance']

    # Component 1: RouteStops I2:I201 — On Time column IF(AND(...)) formulas (0.25 pts)
    # Must FAIL on initial (all None) and PASS on golden (formulas present)
    try:
        i_formula_count = 0
        i_total = 200
        for r in range(2, 202):
            val = ws_rs.cell(row=r, column=9).value
            if val is not None and isinstance(val, str) and val.strip().startswith('='):
                norm = normalize_formula(val)
                # Check for IF(AND(...H...F...H...G...)) pattern
                if 'IF(' in norm and 'AND(' in norm and 'H' in norm and 'F' in norm and 'G' in norm:
                    i_formula_count += 1
        ratio = i_formula_count / i_total
        if i_formula_count >= 190:  # Allow minor tolerance
            print(f"PASS: Component 1 — {i_formula_count}/{i_total} rows in I2:I201 have IF(AND) on-time formula (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — only {i_formula_count}/{i_total} rows in I2:I201 have correct IF(AND) on-time formula (need >=190)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: RouteStops J2:J201 — Late Minutes IF formulas (0.20 pts)
    # Must FAIL on initial (all None) and PASS on golden (formulas present)
    try:
        j_formula_count = 0
        j_total = 200
        for r in range(2, 202):
            val = ws_rs.cell(row=r, column=10).value
            if val is not None and isinstance(val, str) and val.strip().startswith('='):
                norm = normalize_formula(val)
                # Check for IF(I...="LATE",...) or IF(I...=Late) + MAX(0,...)
                if 'IF(' in norm and 'I' in norm and 'LATE' in norm and 'MAX(' in norm:
                    j_formula_count += 1
        if j_formula_count >= 190:
            print(f"PASS: Component 2 — {j_formula_count}/{j_total} rows in J2:J201 have IF(Late)/MAX late minutes formula (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 — only {j_formula_count}/{j_total} rows in J2:J201 have correct late minutes formula (need >=190)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: DriverPerformance B2:B9 — COUNTIF formulas (0.15 pts)
    # Must FAIL on initial (all None) and PASS on golden (COUNTIF formulas)
    try:
        b_formula_count = 0
        for r in range(2, 10):
            val = ws_dp.cell(row=r, column=2).value
            if val is not None and isinstance(val, str) and val.strip().startswith('='):
                norm = normalize_formula(val)
                if 'COUNTIF(' in norm and 'ROUTESTOPS' in norm and 'B' in norm:
                    b_formula_count += 1
        if b_formula_count >= 7:  # Allow 1 miss out of 8
            print(f"PASS: Component 3 — {b_formula_count}/8 rows in B2:B9 have COUNTIF(RouteStops!B) formula (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 — only {b_formula_count}/8 rows in B2:B9 have COUNTIF formula referencing RouteStops!B")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: DriverPerformance C2:C9 — COUNTIFS formulas (0.15 pts)
    # Must FAIL on initial (all None) and PASS on golden (COUNTIFS with On Time condition)
    try:
        c_formula_count = 0
        for r in range(2, 10):
            val = ws_dp.cell(row=r, column=3).value
            if val is not None and isinstance(val, str) and val.strip().startswith('='):
                norm = normalize_formula(val)
                if 'COUNTIFS(' in norm and 'ROUTESTOPS' in norm and 'ONTIME' in norm.replace(' ', '').replace('"', '').replace("'", ""):
                    c_formula_count += 1
        # Fallback: check for COUNTIFS with RouteStops reference and "I" column
        if c_formula_count < 7:
            c_formula_count2 = 0
            for r in range(2, 10):
                val = ws_dp.cell(row=r, column=3).value
                if val is not None and isinstance(val, str) and val.strip().startswith('='):
                    norm = normalize_formula(val)
                    if 'COUNTIFS(' in norm and 'ROUTESTOPS' in norm and '$I' in norm.upper():
                        c_formula_count2 += 1
            c_formula_count = max(c_formula_count, c_formula_count2)
        if c_formula_count >= 7:
            print(f"PASS: Component 4 — {c_formula_count}/8 rows in C2:C9 have COUNTIFS formula with On Time condition (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — only {c_formula_count}/8 rows in C2:C9 have COUNTIFS formula with On Time condition")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: DriverPerformance D2:D9 — On Time % = C/B formulas with percentage format (0.10 pts)
    # Must FAIL on initial (all None) and PASS on golden (C/B formulas + format)
    try:
        d_formula_count = 0
        d_format_count = 0
        for r in range(2, 10):
            cell = ws_dp.cell(row=r, column=4)
            val = cell.value
            if val is not None and isinstance(val, str) and val.strip().startswith('='):
                norm = normalize_formula(val)
                # Check for =Cx/Bx formula pattern
                col_letter = str(r)  # row number used in formula
                if re.search(r'C\d+/B\d+', norm):
                    d_formula_count += 1
            # Check percentage number format
            if cell.number_format and '%' in cell.number_format:
                d_format_count += 1
        if d_formula_count >= 7 and d_format_count >= 6:
            print(f"PASS: Component 5 — {d_formula_count}/8 D rows have C/B formula, {d_format_count}/8 have % format (0.10 pts)")
            total_score += 0.10
        elif d_formula_count >= 7:
            print(f"PARTIAL: Component 5 — {d_formula_count}/8 D rows have C/B formula (no % format), awarding 0.05 pts")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — only {d_formula_count}/8 rows in D2:D9 have C/B on-time % formula (need >=7)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: DriverPerformance E2:E9 — AVERAGEIF formulas for avg late minutes (0.10 pts)
    # Must FAIL on initial (all None) and PASS on golden
    try:
        e_formula_count = 0
        for r in range(2, 10):
            val = ws_dp.cell(row=r, column=5).value
            if val is not None and isinstance(val, str) and val.strip().startswith('='):
                norm = normalize_formula(val)
                if 'AVERAGEIF(' in norm and 'ROUTESTOPS' in norm and '$J' in norm.upper():
                    e_formula_count += 1
        # Fallback: looser check for AVERAGEIF with RouteStops reference
        if e_formula_count < 7:
            e_formula_count2 = 0
            for r in range(2, 10):
                val = ws_dp.cell(row=r, column=5).value
                if val is not None and isinstance(val, str) and val.strip().startswith('='):
                    norm = normalize_formula(val)
                    if 'AVERAGEIF(' in norm and 'ROUTESTOPS' in norm:
                        e_formula_count2 += 1
            e_formula_count = max(e_formula_count, e_formula_count2)
        if e_formula_count >= 7:
            print(f"PASS: Component 6 — {e_formula_count}/8 rows in E2:E9 have AVERAGEIF late minutes formula (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 6 — only {e_formula_count}/8 rows in E2:E9 have AVERAGEIF formula (need >=7)")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: Conditional formatting on D2:D9 — red if <90%, green if >=95% (0.05 pts)
    # Must FAIL on initial (no CF) and PASS on golden (CF present with both rules)
    try:
        cf_rules = ws_dp.conditional_formatting
        red_rule_count = 0
        green_rule_count = 0
        for cf in cf_rules:
            cf_str = str(cf)
            # Check if this CF applies to column D
            if 'D' in cf_str:
                for rule in cf.rules:
                    if rule.type == 'cellIs':
                        # Red rule: lessThan 0.9
                        if rule.operator == 'lessThan':
                            try:
                                formula_val = float(rule.formula[0]) if rule.formula else None
                                if formula_val is not None and abs(formula_val - 0.9) < 0.05:
                                    # Verify fill color is red-ish
                                    if rule.dxf and rule.dxf.fill:
                                        fill_color = rule.dxf.fill.fgColor.rgb
                                        # Red color check (FFFF0000 or similar)
                                        if 'FF0000' in fill_color.upper() or fill_color.upper().endswith('FF0000'):
                                            red_rule_count += 1
                            except Exception:
                                pass
                        # Green rule: greaterThanOrEqual 0.95
                        if rule.operator in ('greaterThanOrEqual', 'greaterThan'):
                            try:
                                formula_val = float(rule.formula[0]) if rule.formula else None
                                if formula_val is not None and formula_val >= 0.93:
                                    green_rule_count += 1
                            except Exception:
                                pass
        if red_rule_count >= 1 and green_rule_count >= 1:
            print(f"PASS: Component 7 — Conditional formatting on D2:D9 has red (<90%) and green (>=95%) rules (0.05 pts)")
            total_score += 0.05
        elif red_rule_count >= 1 or green_rule_count >= 1:
            print(f"PARTIAL: Component 7 — Only partial conditional formatting found (red={red_rule_count}, green={green_rule_count}), awarding 0.025 pts")
            total_score += 0.025
        else:
            print(f"FAIL: Component 7 — No conditional formatting found on D column for <90% red / >=95% green rules")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
