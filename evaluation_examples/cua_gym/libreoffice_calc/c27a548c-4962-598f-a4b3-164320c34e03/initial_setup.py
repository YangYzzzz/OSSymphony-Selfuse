"""
Initial Setup: Multi-drop delivery route log for courier fleet
Task ID: calc_ops_fleet_route_log_070
Domain: libreoffice_calc

Creates a spreadsheet with:
  - Sheet 'RouteStops': 200 stop records with Route ID, Driver, Date, Stop #,
    Customer, Promised Window Start/End, Actual Arrival. Columns I (On Time)
    and J (Late Minutes) are intentionally EMPTY for the agent to fill.
  - Sheet 'DriverPerformance': 8 driver names in column A; columns B-E
    intentionally EMPTY for the agent to fill.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date, datetime, timedelta
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_fleet_route_log_070'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

# Seed for reproducibility
random.seed(42)

DRIVERS = [
    'Marcus Rivera',
    'Leah Thornton',
    'Dimitri Okafor',
    'Priya Nair',
    'Callum Fitzgerald',
    'Yolanda Mestre',
    'Seth Andersen',
    'Renata Kowalski',
]

CUSTOMERS = [
    'Brightfield Bakery', 'Summit Hardware', 'Coastal Pharma', 'Nexus Fitness',
    'Greenway Grocers', 'TechEdge Solutions', 'Harbor Books', 'Sunrise Clinic',
    'Ironworks Gym', 'Maple Lane Cafe', 'Blue Ridge Auto', 'Silver Pines Hotel',
    'Dockside Supplies', 'Pinnacle Apparel', 'Cedar Grove School', 'Mountain View Labs',
    'Riverside Florist', 'Highgate Office Park', 'Lakefront Marina', 'Westfall Pharmacy',
]

# Route ID prefixes per driver
ROUTE_PREFIXES = {d: f'RT-{i+1:02d}' for i, d in enumerate(DRIVERS)}


def random_time(base_hour, base_minute, jitter_minutes=0):
    """Return a datetime.time with optional random jitter."""
    total_min = base_hour * 60 + base_minute + random.randint(-jitter_minutes, jitter_minutes)
    total_min = max(360, min(1140, total_min))  # keep between 6:00 and 19:00
    return datetime(2025, 1, 1, total_min // 60, total_min % 60)


def build_records():
    records = []
    # 200 stops spread across 8 drivers (roughly 25 each)
    # Routes are dated across 4 weeks in March 2025
    base_date = date(2025, 3, 3)  # Monday

    stop_counter = {d: 0 for d in DRIVERS}
    route_counter = {d: 1 for d in DRIVERS}
    current_route = {d: None for d in DRIVERS}
    route_stop_num = {d: 1 for d in DRIVERS}

    for i in range(200):
        driver = DRIVERS[i % len(DRIVERS)]
        stop_counter[driver] += 1

        # Advance route every 3-4 stops per driver
        if stop_counter[driver] % random.choice([3, 4]) == 1:
            route_counter[driver] += 1
            route_stop_num[driver] = 1
            week_offset = (stop_counter[driver] // 25) % 4
            day_offset = random.randint(0, 4)
            current_route[driver] = {
                'date': base_date + timedelta(weeks=week_offset, days=day_offset),
                'route_id': f'{ROUTE_PREFIXES[driver]}-{route_counter[driver]:03d}'
            }
        else:
            route_stop_num[driver] += 1

        if current_route[driver] is None:
            current_route[driver] = {
                'date': base_date,
                'route_id': f'{ROUTE_PREFIXES[driver]}-001'
            }

        route_id = current_route[driver]['route_id']
        stop_date = current_route[driver]['date']
        stop_num = route_stop_num[driver]

        # Promised window: 1-hour window between 08:00 and 17:00
        win_start_min = random.randint(8 * 60, 16 * 60)
        win_end_min = win_start_min + 60  # 1-hour window

        # Actual arrival: mostly on time, some late
        if random.random() < 0.78:
            # On time: within window
            actual_min = random.randint(win_start_min, win_end_min)
        else:
            # Late: arrive 5-45 min after window end
            actual_min = win_end_min + random.randint(5, 45)

        actual_min = min(actual_min, 19 * 60)  # cap at 19:00

        win_start = datetime(stop_date.year, stop_date.month, stop_date.day,
                             win_start_min // 60, win_start_min % 60)
        win_end = datetime(stop_date.year, stop_date.month, stop_date.day,
                           win_end_min // 60, win_end_min % 60)
        actual = datetime(stop_date.year, stop_date.month, stop_date.day,
                          actual_min // 60, actual_min % 60)

        customer = random.choice(CUSTOMERS)

        records.append([
            route_id,        # A: Route ID
            driver,          # B: Driver
            stop_date,       # C: Date
            stop_num,        # D: Stop #
            customer,        # E: Customer
            win_start,       # F: Promised Window Start
            win_end,         # G: Promised Window End
            actual,          # H: Actual Arrival
            # I: On Time — EMPTY (agent must fill)
            # J: Late Minutes — EMPTY (agent must fill)
        ])

    return records


def create_initial():
    wb = openpyxl.Workbook()

    # ── Sheet 1: RouteStops ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'RouteStops'

    headers = [
        'Route ID', 'Driver', 'Date', 'Stop #', 'Customer',
        'Promised Window Start', 'Promised Window End', 'Actual Arrival',
        'On Time', 'Late Minutes'
    ]
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF')

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set column widths
    col_widths = [14, 18, 12, 8, 22, 22, 22, 22, 10, 14]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws1.row_dimensions[1].height = 20

    # Write data rows (A-H only; I and J intentionally empty)
    records = build_records()
    time_fmt = 'yyyy-mm-dd hh:mm'
    date_fmt = 'yyyy-mm-dd'

    for r, row_data in enumerate(records, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            if c == 3:  # Date column
                cell.number_format = date_fmt
            elif c in (6, 7, 8):  # datetime columns
                cell.number_format = time_fmt

    # Freeze header row
    ws1.freeze_panes = 'A2'

    # ── Sheet 2: DriverPerformance ───────────────────────────────────────────
    ws2 = wb.create_sheet('DriverPerformance')

    dp_headers = ['Driver', 'Total Stops', 'On Time Stops', 'On Time %', 'Avg Late Minutes']
    dp_fill = PatternFill(start_color='FF70AD47', end_color='FF70AD47', fill_type='solid')
    dp_font = Font(bold=True, color='FFFFFFFF')

    for col, h in enumerate(dp_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = dp_fill
        cell.font = dp_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    dp_col_widths = [20, 14, 16, 12, 18]
    for i, w in enumerate(dp_col_widths, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws2.row_dimensions[1].height = 20

    # Write driver names (column A only; B-E intentionally empty)
    for r, driver in enumerate(DRIVERS, 2):
        ws2.cell(row=r, column=1, value=driver)

    # Freeze header row
    ws2.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  RouteStops: {len(records)} rows of data (columns A-H filled, I-J empty)')
    print(f'  DriverPerformance: {len(DRIVERS)} drivers in column A (columns B-E empty)')


create_initial()
