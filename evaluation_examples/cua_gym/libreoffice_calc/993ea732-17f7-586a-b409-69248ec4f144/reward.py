"""
Reward Script: Project milestone tracker in LibreOffice Calc
Task ID: calc_grs_023
Domain: libreoffice_calc
Scoring:
  Component 1: Days Variance column exists with formulas (0.25)
  Component 2: Status dropdown data validation (0.20)
  Component 3: Phase dropdown data validation (0.15)
  Component 4: Conditional formatting on Status column (0.25)
  Component 5: Freeze panes on first two columns and top row (0.15)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_023'


def persist_app_state(domain: str):
    """Attempt to save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get the first (and only expected) sheet
    ws = wb.worksheets[0]

    # ---------------------------------------------------------------
    # Component 1: Days Variance column with date-subtraction formulas (0.25 pts)
    # The task asks for a "Days Variance" column that calculates Actual End - Planned End.
    # Initial file has only 10 columns; golden adds column 11 "Days Variance" with formulas.
    # ---------------------------------------------------------------
    try:
        # Check that column 11 (K) has a header related to "Days Variance"
        header_k = ws.cell(row=1, column=11).value
        has_variance_header = (
            header_k is not None
            and "variance" in str(header_k).lower()
        )

        # Count rows with date-subtraction formulas in col 11 (K)
        # Formulas should reference col H (Actual End) and col F (Planned End)
        formula_count = 0
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=11).value
            if val is not None and isinstance(val, str) and '=' in val:
                val_upper = val.upper().replace(" ", "")
                # Check it references H and F columns (date subtraction)
                if 'H' in val_upper and 'F' in val_upper:
                    formula_count += 1

        if has_variance_header and formula_count >= 10:
            print(f"PASS: Component 1 — Days Variance column present with header '{header_k}' and {formula_count} formulas (0.25 pts)")
            total_score += 0.25
        elif has_variance_header and formula_count >= 5:
            print(f"PARTIAL: Component 1 — Header present, but only {formula_count} formula rows (0.15 pts)")
            total_score += 0.15
        elif has_variance_header:
            print(f"PARTIAL: Component 1 — Header present but only {formula_count} formulas (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 1 — No Days Variance column found. Col K header: {header_k}, formulas: {formula_count}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ---------------------------------------------------------------
    # Component 2: Status dropdown data validation (0.20 pts)
    # Task requires: Not Started, In Progress, Completed, Delayed, Cancelled
    # ---------------------------------------------------------------
    try:
        status_dv_score = 0.0
        if ws.data_validations and ws.data_validations.dataValidation:
            for dv in ws.data_validations.dataValidation:
                if dv.type == "list" and dv.formula1:
                    formula_lower = dv.formula1.lower()
                    required_statuses = ["not started", "in progress", "completed", "delayed", "cancelled"]
                    matched = sum(1 for s in required_statuses if s in formula_lower)
                    if matched >= 4 and 'I' in str(dv.sqref).upper():
                        status_dv_score = 0.20
                        print(f"PASS: Component 2 — Status dropdown found with {matched}/5 options, applied to {dv.sqref} (0.20 pts)")
                        break

        if status_dv_score > 0:
            total_score += status_dv_score
        else:
            print(f"FAIL: Component 2 — No Status dropdown data validation found on column I")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---------------------------------------------------------------
    # Component 3: Phase dropdown data validation (0.15 pts)
    # Task requires 4 phase options
    # ---------------------------------------------------------------
    try:
        phase_dv_score = 0.0
        if ws.data_validations and ws.data_validations.dataValidation:
            for dv in ws.data_validations.dataValidation:
                if dv.type == "list" and dv.formula1:
                    sqref_str = str(dv.sqref)
                    if 'C' in sqref_str.upper():
                        options = dv.formula1.strip('"').split(',')
                        if len(options) >= 4:
                            phase_dv_score = 0.15
                            print(f"PASS: Component 3 — Phase dropdown found with {len(options)} options: {dv.formula1}, applied to {sqref_str} (0.15 pts)")
                            break

        if phase_dv_score > 0:
            total_score += phase_dv_score
        else:
            print(f"FAIL: Component 3 — No Phase dropdown data validation with 4+ options found on column C")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ---------------------------------------------------------------
    # Component 4: Conditional formatting on Status column (0.25 pts)
    # Task requires: green for Completed, blue for In Progress, red for Delayed, gray for Not Started
    # We check that there are at least 3 conditional formatting rules on column I range
    # ---------------------------------------------------------------
    try:
        cf_rules_on_status = 0
        status_keywords_found = set()

        for cf in ws.conditional_formatting:
            cf_range_str = str(cf)
            # Check if the range includes column I
            if 'I' in cf_range_str.upper():
                for rule in cf.rules:
                    if hasattr(rule, 'formula') and rule.formula:
                        for f in rule.formula:
                            f_lower = f.lower().strip('"')
                            for kw in ["completed", "in progress", "delayed", "not started"]:
                                if kw in f_lower:
                                    status_keywords_found.add(kw)
                                    cf_rules_on_status += 1

        if cf_rules_on_status >= 4 and len(status_keywords_found) >= 4:
            print(f"PASS: Component 4 — {cf_rules_on_status} conditional formatting rules on Status column covering {status_keywords_found} (0.25 pts)")
            total_score += 0.25
        elif cf_rules_on_status >= 3 and len(status_keywords_found) >= 3:
            print(f"PARTIAL: Component 4 — {cf_rules_on_status} rules covering {status_keywords_found} (0.15 pts)")
            total_score += 0.15
        elif cf_rules_on_status >= 1:
            print(f"PARTIAL: Component 4 — Only {cf_rules_on_status} rules found covering {status_keywords_found} (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4 — No conditional formatting rules found on Status column")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ---------------------------------------------------------------
    # Component 5: Freeze panes — first two columns and top row (0.15 pts)
    # Task says: "Freeze the first two columns and top row"
    # This means freeze_panes should be "C2" (columns A-B and row 1 frozen)
    # ---------------------------------------------------------------
    try:
        fp = ws.freeze_panes
        if fp is not None and str(fp).upper() == "C2":
            print(f"PASS: Component 5 — Freeze panes set to {fp} (first 2 cols + top row) (0.15 pts)")
            total_score += 0.15
        elif fp is not None:
            # Partial credit if some freeze is set but not exactly C2
            print(f"PARTIAL: Component 5 — Freeze panes set to {fp}, expected C2 (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — No freeze panes set (expected C2)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
