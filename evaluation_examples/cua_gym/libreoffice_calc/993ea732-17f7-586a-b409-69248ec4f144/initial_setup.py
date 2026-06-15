"""
Initial Setup: Project milestone tracker for ERP implementation
Task ID: calc_grs_023
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_023'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Milestones"

    # --- Headers (A-J only; K "Days Variance" is part of the task) ---
    headers = [
        "Milestone #", "Milestone Name", "Phase", "Responsible Team",
        "Planned Start", "Planned End", "Actual Start", "Actual End",
        "Status", "Notes"
    ]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Column widths ---
    col_widths = {
        "A": 14, "B": 35, "C": 22, "D": 22,
        "E": 14, "F": 14, "G": 14, "H": 14,
        "I": 14, "J": 30
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Row height for header
    ws.row_dimensions[1].height = 30

    # --- 15 Milestones data ---
    # Phases: Planning, Design, Development, Testing (4 phases for dropdown)
    # Status values used: Not Started, In Progress, Completed, Delayed
    project_start = date(2025, 7, 1)

    milestones = [
        [1, "Project Kickoff & Charter Approval", "Planning", "PMO",
         date(2025, 7, 1), date(2025, 7, 15), date(2025, 7, 1), date(2025, 7, 14),
         "Completed", "All stakeholders aligned on scope and timeline"],
        [2, "Requirements Gathering Workshop", "Planning", "Business Analysis",
         date(2025, 7, 16), date(2025, 8, 5), date(2025, 7, 16), date(2025, 8, 8),
         "Completed", "3 additional sessions needed for finance module"],
        [3, "Current State Process Documentation", "Planning", "Business Analysis",
         date(2025, 8, 6), date(2025, 8, 25), date(2025, 8, 10), date(2025, 8, 28),
         "Completed", "Documented 47 business processes across 6 departments"],
        [4, "Solution Architecture Design", "Design", "Technical Architecture",
         date(2025, 8, 26), date(2025, 9, 20), date(2025, 8, 29), date(2025, 9, 25),
         "Completed", "Hybrid cloud deployment model selected"],
        [5, "Data Migration Strategy", "Design", "Data Engineering",
         date(2025, 9, 15), date(2025, 10, 5), date(2025, 9, 18), date(2025, 10, 12),
         "Completed", "Legacy system has 2.3M records requiring cleansing"],
        [6, "Integration Architecture Blueprint", "Design", "Technical Architecture",
         date(2025, 10, 1), date(2025, 10, 20), date(2025, 10, 3), date(2025, 10, 18),
         "Completed", "14 system integrations mapped and prioritized"],
        [7, "Core Finance Module Configuration", "Development", "Finance Team",
         date(2025, 10, 21), date(2025, 11, 25), date(2025, 10, 22), date(2025, 12, 2),
         "Delayed", "Chart of accounts restructuring took longer than planned"],
        [8, "HR & Payroll Module Setup", "Development", "HR Systems",
         date(2025, 11, 1), date(2025, 12, 10), date(2025, 11, 5), date(2025, 12, 15),
         "Delayed", "Payroll tax rules complexity underestimated"],
        [9, "Supply Chain Module Configuration", "Development", "Operations",
         date(2025, 11, 15), date(2025, 12, 20), date(2025, 11, 18), None,
         "In Progress", "Warehouse management workflow under review"],
        [10, "Data Migration - Phase 1 (Master Data)", "Development", "Data Engineering",
         date(2025, 12, 1), date(2025, 12, 22), date(2025, 12, 5), None,
         "In Progress", "Customer and vendor master data being validated"],
        [11, "Unit Testing - Finance Module", "Testing", "QA Team",
         date(2025, 12, 15), date(2026, 1, 10), None, None,
         "Not Started", "Test scripts drafted, awaiting module completion"],
        [12, "Unit Testing - HR Module", "Testing", "QA Team",
         date(2026, 1, 5), date(2026, 1, 25), None, None,
         "Not Started", "Dependent on milestone 8 completion"],
        [13, "Integration Testing", "Testing", "QA Team",
         date(2026, 1, 20), date(2026, 2, 15), None, None,
         "Not Started", "End-to-end scenarios covering all 14 integrations"],
        [14, "User Acceptance Testing", "Testing", "All Departments",
         date(2026, 2, 10), date(2026, 3, 5), None, None,
         "Not Started", "150 business users identified for UAT participation"],
        [15, "Go-Live & Hypercare Support", "Testing", "PMO",
         date(2026, 3, 15), date(2026, 3, 31), None, None,
         "Not Started", "24/7 support team on standby for first 2 weeks"],
    ]

    date_format = 'yyyy-mm-dd'
    data_align = Alignment(vertical="center")

    for r, row_data in enumerate(milestones, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            cell.alignment = data_align
            # Apply date format to date columns (E, F, G, H = cols 5-8)
            if c in (5, 6, 7, 8) and isinstance(val, date):
                cell.number_format = date_format

    # NO data validation (task asks agent to add dropdowns)
    # NO conditional formatting (task asks agent to add it)
    # NO freeze panes (task asks agent to freeze)
    # NO Days Variance column K (task asks agent to create it)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
