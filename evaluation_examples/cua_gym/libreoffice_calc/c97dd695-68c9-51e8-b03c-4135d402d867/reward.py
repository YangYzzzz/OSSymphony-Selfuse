"""
Reward Script: Fill profit margin formula down column E and create concatenation labels in column F
Task ID: osworld_calc_formula_pattern_concat_003
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5 pts): Column E filled with margin formulas for ALL product rows (E3:E12)
  Component 2 (0.3 pts): Column F header "Label" exists and has concatenation formulas for rows 2-12
  Component 3 (0.2 pts): Column F formulas use TEXT() function with "0.00" format pattern
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_formula_pattern_concat_003'

def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Initial state: E2 has formula '=(D2-C2)/D2*100', E3:E12 are None, no column F.
    Golden state:  E2:E12 all have margin formulas, F1='Label', F2:F12 have concat formulas.
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # --- Precondition gate: expected sheet exists ---
    if 'Products' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Products' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Products']

    # --- Component 1: Column E filled for ALL product rows 2-12 (0.5 pts) ---
    # Initial state: only E2 has a formula, E3:E12 are None.
    # Task requires filling E3:E12 with margin formulas.
    try:
        data_rows = list(range(2, 13))  # rows 2 through 12 (11 product rows)
        e_filled_count = 0
        e_formula_count = 0
        e_missing = []

        for row in data_rows:
            val = ws.cell(row=row, column=5).value  # column E = column 5
            if val is not None:
                e_filled_count += 1
                # Check it looks like the margin formula pattern =(D#-C#)/D#*100
                if isinstance(val, str) and val.startswith('=') and 'D' in val and 'C' in val:
                    e_formula_count += 1
            else:
                e_missing.append(row)

        # We need E3:E12 to be filled (E2 was already filled in initial state).
        # Score based on how many of E3:E12 are filled with formulas.
        rows_to_fill = list(range(3, 13))  # E3 through E12 (10 rows that need filling)
        newly_filled = sum(
            1 for r in rows_to_fill
            if ws.cell(row=r, column=5).value is not None
        )
        newly_with_formula = sum(
            1 for r in rows_to_fill
            if ws.cell(row=r, column=5).value is not None
            and isinstance(ws.cell(row=r, column=5).value, str)
            and ws.cell(row=r, column=5).value.startswith('=')
        )

        if newly_with_formula == 10:
            print(f"PASS: Component 1 — All 10 rows E3:E12 filled with margin formulas (0.5 pts)")
            total_score += 0.5
        elif newly_with_formula >= 5:
            partial = round(0.5 * newly_with_formula / 10, 2)
            print(f"PARTIAL: Component 1 — {newly_with_formula}/10 rows in E3:E12 have margin formulas ({partial} pts)")
            total_score += partial
        elif newly_filled == 10:
            # Filled but not with formulas (e.g., values instead of formulas)
            print(f"PARTIAL: Component 1 — E3:E12 filled with values but not formulas (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 1 — Only {newly_with_formula}/10 rows in E3:E12 have margin formulas (missing rows: {e_missing})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # --- Component 2: Column F header and concatenation formulas for rows 2-12 (0.3 pts) ---
    # Initial state: No column F at all (max_col = 5).
    # Task requires: F1 = "Label" header, F2:F12 = concat formulas.
    try:
        f1_value = ws.cell(row=1, column=6).value  # F1

        # Count how many F rows have formula strings
        f_formula_count = 0
        f_missing = []
        for row in range(2, 13):
            val = ws.cell(row=row, column=6).value
            if val is not None and isinstance(val, str) and val.startswith('='):
                f_formula_count += 1
            else:
                f_missing.append(row)

        has_f_header = f1_value is not None and str(f1_value).strip() != ''

        if f_formula_count == 11 and has_f_header:
            print(f"PASS: Component 2 — F1 header='{f1_value}', all 11 rows F2:F12 have concat formulas (0.3 pts)")
            total_score += 0.3
        elif f_formula_count == 11 and not has_f_header:
            print(f"PARTIAL: Component 2 — F2:F12 all have formulas but F1 header missing (0.2 pts)")
            total_score += 0.2
        elif f_formula_count >= 6:
            partial = round(0.3 * f_formula_count / 11, 2)
            print(f"PARTIAL: Component 2 — {f_formula_count}/11 rows in F2:F12 have concat formulas ({partial} pts)")
            total_score += partial
        elif f_formula_count > 0:
            print(f"PARTIAL: Component 2 — Only {f_formula_count}/11 rows in F2:F12 have concat formulas (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 2 — No column F formulas found (F1='{f1_value}', missing: {f_missing})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # --- Component 3: Column F formulas use TEXT() function with "0.00" format (0.2 pts) ---
    # Verifies that the concatenation properly uses TEXT(E#,"0.00") for percentage formatting.
    # The initial state has no column F at all, so this necessarily fails on initial.
    try:
        text_format_count = 0
        concat_pattern_count = 0
        for row in range(2, 13):
            val = ws.cell(row=row, column=6).value
            if val is not None and isinstance(val, str) and val.startswith('='):
                val_upper = val.upper()
                # Check for TEXT function with 0.00 format
                if 'TEXT(' in val_upper and '0.00' in val:
                    text_format_count += 1
                # Check for concatenation pattern: product name & category & label
                # Pattern: A# & "(" & B# & ")" and some % sign reference
                if re.search(r'A\d+.*B\d+', val) and '%' in val:
                    concat_pattern_count += 1

        if text_format_count == 11:
            print(f"PASS: Component 3 — All 11 F formulas use TEXT() with '0.00' format (0.2 pts)")
            total_score += 0.2
        elif text_format_count >= 6:
            partial = round(0.2 * text_format_count / 11, 2)
            print(f"PARTIAL: Component 3 — {text_format_count}/11 F formulas use TEXT() with '0.00' format ({partial} pts)")
            total_score += partial
        elif concat_pattern_count == 11:
            # Has concat pattern but no TEXT format
            print(f"PARTIAL: Component 3 — Concat pattern found but TEXT(,'0.00') missing (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 3 — TEXT() with '0.00' not found in F formulas (text_format={text_format_count}, concat={concat_pattern_count})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in a given env
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
