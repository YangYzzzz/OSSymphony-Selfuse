"""
Initial Setup: Product profit margin and label spreadsheet
Task ID: osworld_calc_formula_pattern_concat_003
Domain: libreoffice_calc

Creates a product table with:
- Column A: Product Name
- Column B: Category
- Column C: Cost
- Column D: Price
- Column E: Margin % — formula ONLY in E2; E3:E12 are empty
- Column F: does NOT exist (no label column yet)

The agent's task is to:
1. Fill the margin formula down column E for all rows
2. Create column F with concatenation formulas like "Laptop (Electronics): 23.50%"
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_formula_pattern_concat_003'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Products sheet ---
    ws = wb.active
    ws.title = 'Products'

    # Headers (row 1)
    headers = ['Product Name', 'Category', 'Cost', 'Price', 'Margin %']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Product data (realistic content): 11 rows (rows 2-12)
    data = [
        ['Laptop Pro 15',      'Electronics',    820.00,  1249.99],
        ['Wireless Headphones','Electronics',     45.00,    89.99],
        ['Ergonomic Chair',    'Furniture',      180.00,   349.95],
        ['Standing Desk',      'Furniture',      310.00,   599.00],
        ['Coffee Maker Deluxe','Appliances',      55.00,   119.99],
        ['Air Purifier 300',   'Appliances',     130.00,   249.99],
        ['Running Shoes X9',   'Sportswear',      42.00,    95.00],
        ['Yoga Mat Premium',   'Sportswear',      12.50,    34.99],
        ['Blender Pro 1000',   'Appliances',      60.00,   134.99],
        ['Desk Lamp LED',      'Office',           18.00,    44.99],
        ['Mechanical Keyboard','Electronics',      85.00,   159.99],
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        ws.cell(row=r, column=2, value=row_data[1])
        ws.cell(row=r, column=3, value=row_data[2])
        ws.cell(row=r, column=4, value=row_data[3])
        # Column E: formula ONLY in row 2; rows 3-12 are intentionally empty
        if r == 2:
            ws.cell(row=r, column=5, value='=(D2-C2)/D2*100')
        # Column F intentionally left blank (no label column yet)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup — open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
