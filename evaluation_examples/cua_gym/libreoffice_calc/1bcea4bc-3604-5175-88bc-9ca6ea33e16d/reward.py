"""
Reward Script: A/B Test Results Analyzer
Task ID: calc_wf_042
Domain: libreoffice_calc
Scoring:
  Component 1: Conversion rate formulas (F, G columns) — 0.20 points
  Component 2: Statistical formulas (H, I, J, K columns) — 0.25 points
  Component 3: Confidence interval formulas (L-O columns) — 0.15 points
  Component 4: Significance column P with IF formula — 0.15 points
  Component 5: Chart exists with correct structure — 0.15 points
  Component 6: Conditional formatting rules — 0.10 points
"""

import os
import openpyxl


WORKDIR = '/home/user'
TASK_ID = 'calc_wf_042'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice document."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            import time
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that 'AB Test' sheet exists
    if 'AB Test' not in wb.sheetnames:
        print("CRITICAL: 'AB Test' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['AB Test']

    # Component 1: Conversion rate formulas in columns F and G (0.20 points)
    # F = Control Rate = C/B, G = Test Rate = E/D
    # These columns do NOT exist in the initial file (only A-E)
    try:
        f_pass = 0
        g_pass = 0
        for row in range(2, 12):  # rows 2-11
            f_val = ws.cell(row=row, column=6).value  # Column F
            g_val = ws.cell(row=row, column=7).value  # Column G
            if f_val and isinstance(f_val, str) and '/' in f_val.upper() and 'C' in f_val.upper() and 'B' in f_val.upper():
                f_pass += 1
            elif f_val and isinstance(f_val, (int, float)):
                # Could be computed value — also acceptable if reasonable
                f_pass += 1
            if g_val and isinstance(g_val, str) and '/' in g_val.upper() and 'E' in g_val.upper() and 'D' in g_val.upper():
                g_pass += 1
            elif g_val and isinstance(g_val, (int, float)):
                g_pass += 1

        # Need at least 8 out of 10 for each to pass
        if f_pass >= 8 and g_pass >= 8:
            print(f"PASS: Component 1 — Conversion rate formulas found (F:{f_pass}/10, G:{g_pass}/10) (0.20 pts)")
            total_score += 0.20
        elif f_pass >= 5 or g_pass >= 5:
            partial = 0.10
            print(f"PARTIAL: Component 1 — Some conversion rate formulas (F:{f_pass}/10, G:{g_pass}/10) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Conversion rate formulas missing (F:{f_pass}/10, G:{g_pass}/10)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Statistical formulas — Pooled Proportion (H), Standard Error (I),
    # Z-Score (J), P-Value (K) — 0.25 points
    try:
        stat_cols_pass = {8: 0, 9: 0, 10: 0, 11: 0}  # H, I, J, K
        col_names = {8: 'Pooled Prop (H)', 9: 'Std Error (I)', 10: 'Z-Score (J)', 11: 'P-Value (K)'}

        for row in range(2, 12):
            for col in [8, 9, 10, 11]:
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    if isinstance(val, str) and val.startswith('='):
                        stat_cols_pass[col] += 1
                    elif isinstance(val, (int, float)):
                        stat_cols_pass[col] += 1

        # Check headers exist for these columns
        h_headers = []
        for col in [8, 9, 10, 11]:
            hdr = ws.cell(row=1, column=col).value
            if hdr is not None:
                h_headers.append(col)

        cols_with_data = sum(1 for c in [8, 9, 10, 11] if stat_cols_pass[c] >= 8)

        if cols_with_data >= 4:
            print(f"PASS: Component 2 — All 4 statistical formula columns present (0.25 pts)")
            for col in [8, 9, 10, 11]:
                print(f"  {col_names[col]}: {stat_cols_pass[col]}/10 rows")
            total_score += 0.25
        elif cols_with_data >= 2:
            partial = 0.15
            print(f"PARTIAL: Component 2 — {cols_with_data}/4 statistical columns present ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Statistical formula columns missing")
            for col in [8, 9, 10, 11]:
                print(f"  {col_names[col]}: {stat_cols_pass[col]}/10 rows")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Confidence interval formulas in columns L-O (0.15 points)
    # L=Control CI Lower, M=Control CI Upper, N=Test CI Lower, O=Test CI Upper
    try:
        ci_cols_pass = {12: 0, 13: 0, 14: 0, 15: 0}  # L, M, N, O
        col_names_ci = {12: 'Ctrl CI Lower (L)', 13: 'Ctrl CI Upper (M)',
                        14: 'Test CI Lower (N)', 15: 'Test CI Upper (O)'}

        for row in range(2, 12):
            for col in [12, 13, 14, 15]:
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    if isinstance(val, str) and val.startswith('='):
                        # Check for 1.96 in CI formulas (z-critical value)
                        if '1.96' in val:
                            ci_cols_pass[col] += 1
                        else:
                            ci_cols_pass[col] += 1  # Still a formula, accept
                    elif isinstance(val, (int, float)):
                        ci_cols_pass[col] += 1

        cols_with_ci = sum(1 for c in [12, 13, 14, 15] if ci_cols_pass[c] >= 8)

        if cols_with_ci >= 4:
            print(f"PASS: Component 3 — All 4 CI columns present (0.15 pts)")
            total_score += 0.15
        elif cols_with_ci >= 2:
            partial = 0.08
            print(f"PARTIAL: Component 3 — {cols_with_ci}/4 CI columns ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — CI columns missing (found {cols_with_ci}/4)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Significance column P with IF formula (0.15 points)
    # P should contain =IF(ABS(J)>1.96,"Yes","No") or similar
    try:
        p_pass = 0
        for row in range(2, 12):
            val = ws.cell(row=row, column=16).value  # Column P
            if val is not None:
                if isinstance(val, str):
                    val_upper = val.upper()
                    if val_upper.startswith('=') and 'IF' in val_upper and '1.96' in val:
                        p_pass += 1
                    elif val_upper in ('YES', 'NO'):
                        # Computed result — acceptable
                        p_pass += 1

        if p_pass >= 8:
            print(f"PASS: Component 4 — Significance column P present ({p_pass}/10 rows) (0.15 pts)")
            total_score += 0.15
        elif p_pass >= 5:
            partial = 0.08
            print(f"PARTIAL: Component 4 — Some significance values ({p_pass}/10 rows) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Significance column P missing or incomplete ({p_pass}/10 rows)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Chart exists with correct structure (0.15 points)
    # Should be a bar chart comparing control vs test conversion rates, with 2 series
    try:
        charts = ws._charts
        if len(charts) >= 1:
            chart = charts[0]
            series_count = len(chart.series)

            # Check it's a bar/column chart
            is_bar = chart.__class__.__name__ == 'BarChart'

            if is_bar and series_count >= 2:
                print(f"PASS: Component 5 — BarChart with {series_count} series found (0.15 pts)")
                total_score += 0.15
            elif series_count >= 2:
                partial = 0.10
                print(f"PARTIAL: Component 5 — Chart found ({chart.__class__.__name__}) with {series_count} series ({partial} pts)")
                total_score += partial
            elif len(charts) >= 1:
                partial = 0.05
                print(f"PARTIAL: Component 5 — Chart found but only {series_count} series ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 5 — Chart type/series mismatch")
        else:
            print(f"FAIL: Component 5 — No charts found in 'AB Test' sheet")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Conditional formatting rules (0.10 points)
    # Should have rules to highlight statistically significant rows (green fill)
    try:
        cf_rules = list(ws.conditional_formatting)
        cf_count = len(cf_rules)

        if cf_count >= 5:
            # Check that at least some rules reference significance / column P
            sig_rules = 0
            for cf in cf_rules:
                for rule in cf.rules:
                    formula_list = getattr(rule, 'formula', [])
                    if formula_list:
                        for f in formula_list:
                            f_upper = str(f).upper()
                            if 'P' in f_upper or 'YES' in f_upper or 'SIGNIFICANT' in f_upper:
                                sig_rules += 1
                                break

            if sig_rules >= 5:
                print(f"PASS: Component 6 — {cf_count} conditional formatting rules, {sig_rules} reference significance (0.10 pts)")
                total_score += 0.10
            elif cf_count >= 5:
                # Has enough CF rules even if we can't confirm they reference P
                print(f"PASS: Component 6 — {cf_count} conditional formatting rules found (0.10 pts)")
                total_score += 0.10
            else:
                partial = 0.05
                print(f"PARTIAL: Component 6 — Some CF rules found ({cf_count}) ({partial} pts)")
                total_score += partial
        elif cf_count >= 1:
            partial = 0.05
            print(f"PARTIAL: Component 6 — {cf_count} conditional formatting rules (need >=5) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 6 — No conditional formatting rules found")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
