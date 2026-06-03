"""
Initial Setup: A/B Test Results Analyzer
Task ID: calc_wf_042
Domain: libreoffice_calc

Creates a workbook with raw A/B test data (10 tests) for control and test groups.
No formulas, no charts, no conditional formatting - those are the task for the agent.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_042'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "AB Test"

    # --- Headers ---
    headers = [
        "Test Name",
        "Control Visitors",
        "Control Conversions",
        "Test Visitors",
        "Test Conversions",
    ]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- A/B Test Data (10 tests) ---
    # Realistic experiment names with meaningful visitor/conversion numbers
    data = [
        ["Homepage Hero Banner",      12450, 387,  12380, 425],
        ["Checkout Button Color",      8920, 214,   8875, 268],
        ["Pricing Page Layout",       15300, 612,  15250, 589],
        ["Email Subject Line",        22100, 1547, 22050, 1654],
        ["Product Image Size",         6780, 203,   6810, 241],
        ["Navigation Menu Style",     18500, 555,  18420, 621],
        ["Landing Page Copy",         10200, 306,  10150, 345],
        ["Search Bar Placement",       9450, 283,   9500, 312],
        ["Mobile CTA Position",       14800, 592,  14750, 634],
        ["Cart Abandonment Popup",    11600, 464,  11550, 439],
    ]

    data_font = Font(name="Calibri", size=11)
    data_align_text = Alignment(horizontal="left", vertical="center")
    data_align_num = Alignment(horizontal="center", vertical="center")

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = thin_border
            if c == 1:
                cell.alignment = data_align_text
            else:
                cell.alignment = data_align_num
                if c in (2, 4):  # visitors columns
                    cell.number_format = '#,##0'
                elif c in (3, 5):  # conversions columns
                    cell.number_format = '#,##0'

    # --- Column Widths ---
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 20

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
