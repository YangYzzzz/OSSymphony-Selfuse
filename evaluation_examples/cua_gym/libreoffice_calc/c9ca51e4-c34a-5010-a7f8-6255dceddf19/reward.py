"""
Reward Script: Calculate running expense total using whole-column SUM reference
Task ID: calc_fmb_sum_entire_col_048
Domain: libreoffice_calc
Scoring:
  Component 1: Cell D1 contains a SUM formula (0.5 pts)
  Component 2: SUM formula uses whole-column reference B:B (0.5 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmb_sum_entire_col_048'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.

    Task: Put =SUM(B:B) in cell D1 of the 'Expenses' sheet.
    The formula should use a whole-column reference (B:B), not a range (B2:B348),
    so that it continues to work as new rows are added.

    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition gate: ensure file can be loaded
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: ensure 'Expenses' sheet exists
    if 'Expenses' not in wb.sheetnames:
        print(f"CRITICAL: Sheet 'Expenses' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Expenses']

    # Component 1: Cell D1 contains a SUM formula (0.5 points)
    # The task requires placing a formula in D1. Initial state has D1=None.
    # Any SUM formula in D1 earns partial credit.
    try:
        d1_value = ws['D1'].value
        if d1_value is not None and isinstance(d1_value, str) and d1_value.strip().upper().startswith('=SUM('):
            print(f"PASS: Component 1 — D1 contains a SUM formula: {repr(d1_value)} (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — Expected D1 to contain =SUM(...), found: {repr(d1_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not check D1: {e}")

    # Component 2: SUM formula uses whole-column reference B:B (0.5 points)
    # The task specifically requires a whole-column reference (B:B) rather than
    # a fixed range (e.g. B2:B348), so the formula scales as data grows.
    # This requires Component 1 to have passed (D1 must already be a SUM formula).
    try:
        d1_value = ws['D1'].value
        if d1_value is not None and isinstance(d1_value, str):
            # Normalize: remove spaces, uppercase for comparison
            normalized = d1_value.replace(' ', '').upper()
            if 'B:B' in normalized:
                print(f"PASS: Component 2 — D1 uses whole-column reference B:B: {repr(d1_value)} (0.5 pts)")
                total_score += 0.5
            else:
                print(f"FAIL: Component 2 — D1 formula does not use whole-column B:B reference. Found: {repr(d1_value)}")
        else:
            print(f"FAIL: Component 2 — D1 is empty or not a formula, cannot check column reference")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not check B:B reference in D1: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
