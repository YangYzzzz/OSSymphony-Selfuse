"""
Initial Setup: Expenses spreadsheet for SUM whole-column formula task
Task ID: calc_fmb_sum_entire_col_048
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_sum_entire_col_048'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Expenses ---
    ws = wb.active
    ws.title = 'Expenses'

    # Row 1: headers
    # B1 = 'Amount (Header)' as specified in task context
    # C1 = 'Total Expenses' as specified in task context
    # D1 = empty (target cell — MUST remain empty in initial)
    ws['A1'] = 'Category'
    ws['B1'] = 'Amount (Header)'
    ws['C1'] = 'Total Expenses'
    ws['D1'] = None  # target cell — must be empty

    # 347 expense data rows (B2:B348) with realistic business expenses
    # Target sum = 1,284,650
    # Strategy: use a mix of realistic expense amounts, adjust last row to hit exact sum

    categories = [
        'Office Supplies', 'Travel & Transportation', 'Meals & Entertainment',
        'Software Licenses', 'Hardware Equipment', 'Utilities',
        'Marketing & Advertising', 'Professional Services', 'Training & Development',
        'Facility Maintenance', 'Telecommunications', 'Insurance',
        'Contractor Fees', 'Research & Development', 'Printing & Stationery',
        'Shipping & Postage', 'Cleaning Services', 'Security Services',
        'IT Support', 'Legal Fees', 'Accounting Services', 'Recruitment',
        'Employee Benefits', 'Subscriptions', 'Cloud Services',
    ]

    vendors = [
        'Staples', 'Delta Airlines', 'Marriott Hotels', 'Adobe Systems',
        'Dell Technologies', 'Pacific Gas & Electric', 'Google Ads',
        'McKinsey & Co', 'Coursera Enterprise', 'HVAC Solutions Inc',
        'AT&T Business', 'Travelers Insurance', 'TechStaff Solutions',
        'MIT Research Group', 'FedEx Office', 'UPS Logistics',
        'CleanCo Services', 'SecureGuard Inc', 'CompuRepair Pro',
        'Baker & McKenzie', 'Deloitte Tax', 'LinkedIn Talent',
        'ADP Benefits', 'Salesforce', 'Amazon Web Services',
    ]

    expense_amounts = [
        245.50, 1820.00, 387.25, 4500.00, 8750.00, 612.30,
        3200.00, 12500.00, 1450.00, 2800.00, 890.00, 3600.00,
        9500.00, 5400.00, 175.80, 420.60, 650.00, 1200.00,
        2100.00, 7500.00, 4800.00, 1900.00, 6300.00, 540.00, 2400.00,
        310.75, 2650.00, 495.00, 3800.00, 11200.00, 760.40,
        4100.00, 15000.00, 1750.00, 3100.00, 940.00, 4200.00,
        10500.00, 6200.00, 230.20, 510.80, 720.00, 1350.00,
        2400.00, 8200.00, 5100.00, 2050.00, 7100.00, 620.00, 2700.00,
        280.90, 2950.00, 445.00, 4100.00, 9800.00, 695.60,
        3600.00, 13800.00, 1600.00, 2900.00, 870.00, 3900.00,
        11000.00, 5700.00, 195.40, 465.30, 700.00, 1280.00,
        2250.00, 7800.00, 5300.00, 2150.00, 6700.00, 580.00, 2550.00,
        265.60, 2750.00, 520.00, 3950.00, 10200.00, 735.80,
        3850.00, 14200.00, 1680.00, 2850.00, 915.00, 4050.00,
        10800.00, 6000.00, 215.70, 490.50, 680.00, 1320.00,
        2350.00, 8000.00, 5200.00, 2100.00, 6900.00, 600.00, 2620.00,
        292.30, 2820.00, 468.00, 3870.00, 10000.00, 705.20,
        3700.00, 14500.00, 1720.00, 2920.00, 880.00, 3750.00,
        10300.00, 5850.00, 205.50, 478.40, 710.00, 1300.00,
        2300.00, 7900.00, 5150.00, 2075.00, 6800.00, 590.00, 2580.00,
        272.40, 2790.00, 482.00, 4020.00, 9900.00, 718.60,
        3750.00, 14800.00, 1695.00, 2880.00, 895.00, 3825.00,
        10550.00, 5925.00, 210.60, 484.90, 695.00, 1310.00,
        2325.00, 7950.00, 5175.00, 2087.50, 6850.00, 595.00, 2595.00,
        285.10, 2810.00, 475.00, 3985.00, 9950.00, 712.40,
        3725.00, 13500.00, 1710.00, 2910.00, 887.50, 3787.50,
        10425.00, 5887.50, 207.55, 481.65, 702.50, 1305.00,
        2312.50, 7925.00, 5162.50, 2081.25, 6825.00, 592.50, 2587.50,
        278.75, 2800.00, 478.50, 4002.50, 9925.00, 715.50,
        3737.50, 14650.00, 1702.50, 2895.00, 891.25, 3806.25,
        10487.50, 5906.25, 208.58, 483.28, 697.50, 1307.50,
        2318.75, 7937.50, 5168.75, 2084.38, 6837.50, 593.75, 2591.25,
        281.93, 2805.00, 476.75, 3993.75, 9937.50, 713.95,
        3731.25, 14325.00, 1706.25, 2902.50, 889.38, 3796.88,
        10456.25, 5896.88, 208.07, 482.47, 699.75, 1306.25,
        2315.63, 7931.25, 5165.63, 2082.82, 6831.25, 593.13, 2589.38,
        280.34, 2802.50, 477.63, 3998.13, 9931.25, 714.73,
        3728.13, 14987.50, 1708.13, 2898.75, 890.32, 3801.57,
        10471.88, 5901.57, 208.33, 482.88, 698.63, 1306.88,
        2317.19, 7934.38, 5167.19, 2083.60, 6834.38, 593.44, 2590.32,
        281.14, 2803.75, 477.19, 3995.95, 9934.38, 714.34,
        3729.69, 14656.25, 1707.19, 2900.63, 889.85, 3799.23,
        10463.07, 5899.23, 208.20, 482.68, 699.19, 1306.57,
        2316.41, 7932.82, 5166.41, 2083.21, 6832.82, 593.29, 2589.85,
        280.74, 2803.13, 477.41, 3997.04, 9932.82, 714.54,
        320.00, 2900.00, 500.00, 4200.00, 9200.00, 640.00,
        3400.00, 14000.00, 1600.00, 3000.00, 880.00, 3500.00,
        10000.00, 5500.00, 200.00, 450.00, 670.00, 1250.00,
        2200.00, 7600.00, 5000.00, 2000.00, 6600.00, 560.00, 2500.00,
        260.00, 2700.00, 460.00, 3800.00, 9700.00, 700.00,
        3600.00, 13200.00, 1600.00, 2800.00, 840.00, 3700.00,
        10000.00, 5600.00, 190.00, 460.00, 660.00, 1240.00,
        2180.00, 7700.00, 5050.00, 2020.00, 6700.00, 570.00, 2530.00,
        270.00, 2750.00, 470.00, 3880.00, 9750.00, 708.00,
        3640.00, 13600.00, 1640.00, 2860.00, 858.00, 3740.00,
        10200.00, 5640.00, 198.00, 464.00, 668.00, 1248.00,
        2216.00, 7740.00, 5040.00, 2028.00, 6660.00, 566.00, 2516.00,
    ]

    # We need exactly 347 rows of data
    # Build amounts list cycling through the expense_amounts array
    row_amounts = []
    for i in range(346):  # first 346 rows
        row_amounts.append(expense_amounts[i % len(expense_amounts)])

    # Calculate the last amount to hit the exact target sum
    TARGET_SUM = 1284650
    current_sum = sum(row_amounts)
    last_amount = TARGET_SUM - current_sum
    row_amounts.append(round(last_amount, 2))

    # Write data rows
    for i, amount in enumerate(row_amounts):
        row = i + 2  # data starts at row 2
        cat = categories[i % len(categories)]
        vendor = vendors[i % len(vendors)]
        ws.cell(row=row, column=1, value=cat)    # A: Category
        ws.cell(row=row, column=2, value=amount)  # B: Amount
        ws.cell(row=row, column=3, value=vendor)  # C: Vendor name

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Expenses')
    print(f'  Header row: A1=Category, B1=Amount (Header), C1=Total Expenses, D1=<empty>')
    print(f'  Data rows: {len(row_amounts)} (B2:B{1 + len(row_amounts)})')
    print(f'  Sum of B column (B2:B{1 + len(row_amounts)}): {sum(row_amounts):,.2f}')
    print(f'  D1 is empty (target cell for SUM formula)')


create_initial()
