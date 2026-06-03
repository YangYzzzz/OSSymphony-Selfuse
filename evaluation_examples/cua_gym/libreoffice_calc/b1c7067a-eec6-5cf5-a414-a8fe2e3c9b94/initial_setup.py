"""
Initial Setup: Patient admission table with ward data across 12 months.
Task ID: osworld_calc_total_row_line_chart_004
Domain: libreoffice_calc

Creates a spreadsheet with hospital ward admission data:
  - Column A: Ward names
  - Columns B-M: Monthly admission counts (Jan-Dec)
  - 8 hospital wards with realistic data
  - NO total row, NO chart (those are the task for the agent)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_total_row_line_chart_004'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patient Admissions"

    # --- Headers ---
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    headers = ["Ward"] + months
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name="Calibri", size=11)
        cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        cell.font = Font(bold=True, name="Calibri", size=11, color="FFFFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # --- Ward admission data (realistic hospital data) ---
    # 8 hospital wards with monthly admission counts
    ward_data = [
        ["Cardiology",     142, 138, 155, 149, 163, 158, 172, 168, 161, 154, 147, 160],
        ["Orthopedics",     98,  92, 105, 111, 118, 103,  96, 102, 108, 115, 121, 109],
        ["Pediatrics",     187, 175, 192, 203, 218, 225, 231, 228, 214, 199, 188, 196],
        ["Neurology",       76,  71,  83,  79,  88,  85,  91,  87,  84,  80,  74,  82],
        ["Oncology",        64,  68,  72,  69,  75,  71,  78,  74,  70,  67,  65,  73],
        ["Emergency",      312, 289, 335, 298, 342, 356, 368, 374, 351, 329, 315, 348],
        ["General Surgery", 124, 118, 131, 127, 139, 133, 141, 136, 129, 124, 119, 135],
        ["Maternity",       89,  82,  95, 101, 107,  98,  93,  99, 104, 110, 116, 103],
    ]

    for r, row_data in enumerate(ward_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 1:
                cell.font = Font(name="Calibri", size=11)
            else:
                cell.alignment = Alignment(horizontal="center")

    # --- Column widths ---
    ws.column_dimensions["A"].width = 18
    for col_letter in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]:
        ws.column_dimensions[col_letter].width = 6

    # --- Freeze the header row ---
    ws.freeze_panes = "B2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
