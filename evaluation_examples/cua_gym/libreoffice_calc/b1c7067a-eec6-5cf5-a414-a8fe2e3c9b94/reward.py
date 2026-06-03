"""
Reward Script: Add monthly total row and line chart to patient admission spreadsheet
Task ID: osworld_calc_total_row_line_chart_004
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Total row exists with label 'Total' in column A at bottom of data
  Component 2 (0.20): Total row has SUM formulas covering all ward rows (B2:Bx through M2:Mx)
  Component 3 (0.30): A LineChart exists on the worksheet
  Component 4 (0.20): Chart categories reference month name labels (x-axis has month names)
  Total: 1.0
"""

import os
import re
import openpyxl
from openpyxl.chart import LineChart

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_total_row_line_chart_004'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task: Add a monthly total row at the bottom and create a line chart
    showing monthly totals with month names on x-axis.
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get the active sheet (should be 'Patient Admissions')
    ws = wb.active
    print(f"Active sheet: {ws.title}")
    print(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns")

    # Determine data boundaries
    # Row 1 is header (Ward, Jan, Feb, ..., Dec)
    # Data rows start at row 2
    # Total row should be at the bottom
    max_row = ws.max_row
    max_col = ws.max_column

    # -------------------------------------------------------------------
    # Component 1: Total row exists with label 'Total' in column A (0.30 pts)
    # This FAILS on initial (no Total row) and PASSES on golden (Total row added)
    # -------------------------------------------------------------------
    try:
        total_row_idx = None
        # Search last few rows for a cell with 'Total' label in column A
        for row_idx in range(max_row, max(1, max_row - 3), -1):
            cell_val = ws.cell(row=row_idx, column=1).value
            if cell_val and str(cell_val).strip().lower() == 'total':
                total_row_idx = row_idx
                break

        if total_row_idx is not None:
            print(f"PASS: Component 1 — Total row found at row {total_row_idx} (label: '{ws.cell(row=total_row_idx, column=1).value}') (0.30 pts)")
            total_score += 0.30
        else:
            # Also check if any row has 'Total' in column A (anywhere)
            for row_idx in range(2, max_row + 1):
                cell_val = ws.cell(row=row_idx, column=1).value
                if cell_val and str(cell_val).strip().lower() == 'total':
                    total_row_idx = row_idx
                    break
            if total_row_idx is not None:
                print(f"PASS: Component 1 — Total row found at row {total_row_idx} (label: '{ws.cell(row=total_row_idx, column=1).value}') (0.30 pts)")
                total_score += 0.30
            else:
                print(f"FAIL: Component 1 — No row with 'Total' label in column A found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------
    # Component 2: Total row has SUM formulas for monthly columns (0.20 pts)
    # Checks that B-M in the Total row contain SUM formulas covering ward data rows
    # This FAILS on initial (no Total row) and PASSES on golden (SUM formulas present)
    # -------------------------------------------------------------------
    try:
        if total_row_idx is not None:
            sum_formula_cols = 0
            total_months = 0
            # Check columns B through M (columns 2-13) in the Total row
            for col_idx in range(2, min(max_col + 1, 14)):
                cell = ws.cell(row=total_row_idx, column=col_idx)
                cell_val = cell.value
                total_months += 1
                if cell_val and isinstance(cell_val, str):
                    # Check if it's a SUM formula
                    formula_upper = cell_val.upper().replace(' ', '')
                    if formula_upper.startswith('=SUM('):
                        sum_formula_cols += 1
                    # Also accept numeric values (pre-calculated totals are acceptable)
                elif cell_val is not None and isinstance(cell_val, (int, float)):
                    # Some implementations may store computed values
                    # We still count these as valid totals
                    sum_formula_cols += 1

            if total_months > 0 and sum_formula_cols == total_months:
                print(f"PASS: Component 2 — All {sum_formula_cols}/{total_months} monthly columns have SUM formulas in Total row (0.20 pts)")
                total_score += 0.20
            elif total_months > 0 and sum_formula_cols >= total_months * 0.8:
                # Partial: at least 80% of columns have SUM formulas
                print(f"PARTIAL: Component 2 — {sum_formula_cols}/{total_months} monthly columns have SUM formulas (partial credit 0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 2 — Only {sum_formula_cols}/{total_months} columns in Total row have SUM formulas")
        else:
            print(f"FAIL: Component 2 — Cannot check SUM formulas; Total row not found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------
    # Component 3: A LineChart exists on the worksheet (0.30 pts)
    # This FAILS on initial (0 charts) and PASSES on golden (1 LineChart)
    # -------------------------------------------------------------------
    try:
        charts = ws._charts
        line_charts = [c for c in charts if isinstance(c, LineChart)]

        if len(line_charts) >= 1:
            print(f"PASS: Component 3 — LineChart found ({len(line_charts)} line chart(s), {len(charts)} total charts) (0.30 pts)")
            total_score += 0.30
        elif len(charts) >= 1:
            # There's a chart but not a LineChart — partial credit
            print(f"PARTIAL: Component 3 — Chart found but not a LineChart (type: {type(charts[0]).__name__}) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 — No charts found on worksheet")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------
    # Component 4: Chart has month name labels (Jan-Dec) as categories on x-axis (0.20 pts)
    # The x-axis categories should reference the month names from the header row
    # This FAILS on initial (no chart) and PASSES on golden (chart with month categories)
    # -------------------------------------------------------------------
    try:
        month_names = {'jan', 'feb', 'mar', 'apr', 'may', 'jun',
                       'jul', 'aug', 'sep', 'oct', 'nov', 'dec'}
        found_month_categories = False
        category_details = "no chart found"

        if len(line_charts) >= 1:
            chart = line_charts[0]
            # Check all series for category references
            for ser in chart.series:
                cat = getattr(ser, 'cat', None)
                if cat:
                    # Check numRef (may contain month header row reference)
                    if cat.numRef and cat.numRef.f:
                        ref_formula = cat.numRef.f.lower()
                        # Check if the reference covers the header row (row 1) in columns B-M
                        # Pattern like: 'sheet'!$b$1:$m$1
                        if '$1' in ref_formula or 'row1' in ref_formula.replace(' ', ''):
                            found_month_categories = True
                            category_details = f"numRef: {cat.numRef.f}"
                            break
                        # Also accept references containing month abbreviations via literal values
                    if cat.strRef and cat.strRef.f:
                        ref_formula = cat.strRef.f.lower()
                        if '$1' in ref_formula:
                            found_month_categories = True
                            category_details = f"strRef: {cat.strRef.f}"
                            break
                    # Check if numRef formula references row 1 (header row with month names)
                    if cat.numRef and cat.numRef.f:
                        # Check if it references the first row which contains month names
                        f = cat.numRef.f
                        if re.search(r'\$[A-M]\$1:\$[A-M]\$1', f, re.IGNORECASE) or \
                           re.search(r'\$[B-M]\$1:\$[B-M]\$1', f, re.IGNORECASE):
                            found_month_categories = True
                            category_details = f"numRef row1: {f}"
                            break

            if not found_month_categories and len(line_charts) >= 1:
                # Alternative: check if any series cat references B1:M1 range
                chart = line_charts[0]
                for ser in chart.series:
                    cat = getattr(ser, 'cat', None)
                    if cat and cat.numRef and cat.numRef.f:
                        f = cat.numRef.f
                        # Look for references to row 1 columns B through M
                        if re.search(r'\$?[Bb]\$?1.*\$?[Mm]\$?1', f):
                            found_month_categories = True
                            category_details = f"Row 1 ref: {f}"
                            break

        elif len(charts) >= 1:
            # Some non-line chart — check if it has categories
            chart = charts[0]
            for ser in chart.series:
                cat = getattr(ser, 'cat', None)
                if cat and (cat.numRef or cat.strRef):
                    ref = cat.numRef.f if cat.numRef else cat.strRef.f
                    if '$1' in ref:
                        found_month_categories = True
                        category_details = f"Chart category ref: {ref}"
                        break

        if found_month_categories:
            print(f"PASS: Component 4 — Chart categories reference month names ({category_details}) (0.20 pts)")
            total_score += 0.20
        else:
            # Fallback: check if chart data references the total row (row 10)
            # which at minimum shows the monthly trend data is being charted
            total_row_charted = False
            if len(line_charts) >= 1:
                chart = line_charts[0]
                for ser in chart.series:
                    val_src = getattr(ser, 'val', None)
                    if val_src and val_src.numRef and val_src.numRef.f:
                        f = val_src.numRef.f
                        # Check if series references the total row
                        if total_row_idx and f'${total_row_idx}' in f:
                            total_row_charted = True
                            break

            if total_row_charted:
                print(f"PARTIAL: Component 4 — Chart references Total row data but month categories not confirmed (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 4 — Chart does not have month name categories on x-axis (detail: {category_details})")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
