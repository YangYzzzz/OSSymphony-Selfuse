"""
Initial Setup: Territory Assignment Form — Dependent Dropdown Task
Task ID: osworld_calc_data_validation_dropdown_007
Domain: libreoffice_calc

Creates a territory assignment spreadsheet with:
- TerritoryAssignment sheet: main form with Territory, Region, Subregion columns
- RegionData sheet: reference sheet listing subregions per region (NO named ranges yet)
- NO data validation on columns B or C (task adds those)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_data_validation_dropdown_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ── Sheet 1: TerritoryAssignment ───────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'TerritoryAssignment'

    # Style helpers
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
    header_fill = PatternFill(start_color='FF2E74B5', end_color='FF2E74B5', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='FF000000')
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['Territory ID', 'Region', 'Subregion', 'Sales Rep', 'Q1 Target ($)', 'Q2 Target ($)', 'Status']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = cell_border

    # Realistic territory data (NO validation — task will add it)
    # Region and Subregion columns are left as plain text (no dropdowns yet)
    data = [
        ['TER-001', 'North',  'Minnesota',      'James Harrington',  125000, 130000, 'Active'],
        ['TER-002', 'South',  'Florida',         'Carmen Rodriguez',  142000, 148000, 'Active'],
        ['TER-003', 'East',   'New York',         'David Park',        198000, 205000, 'Active'],
        ['TER-004', 'West',   'California',       'Rachel Kim',        210000, 220000, 'Active'],
        ['TER-005', 'North',  'Wisconsin',        'Tom Mackenzie',      98000, 102000, 'Active'],
        ['TER-006', 'South',  'Georgia',          'Latoya Williams',   115000, 119000, 'Review'],
        ['TER-007', 'East',   'Massachusetts',    'Brian Sullivan',    175000, 182000, 'Active'],
        ['TER-008', 'West',   'Washington',       'Diana Chen',        163000, 170000, 'Active'],
        ['TER-009', 'North',  'Michigan',         'Carlos Mendez',      88000,  94000, 'Active'],
        ['TER-010', 'South',  'Texas',            'Alicia Moore',      187000, 195000, 'Active'],
        ['TER-011', 'East',   'Pennsylvania',     'Kevin Okafor',      144000, 151000, 'Review'],
        ['TER-012', 'West',   'Oregon',           'Priya Sharma',      129000, 135000, 'Active'],
        ['TER-013', 'North',  'Minnesota',        'Sandra Bell',        92000,  97000, 'Inactive'],
        ['TER-014', 'South',  'Alabama',          'Marcus Johnson',     79000,  84000, 'Active'],
        ['TER-015', 'East',   'New Jersey',       'Sarah Chen',        156000, 163000, 'Active'],
    ]

    row_fill_even = PatternFill(start_color='FFDCE6F1', end_color='FFDCE6F1', fill_type='solid')
    data_align = Alignment(horizontal='left', vertical='center')

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.border = cell_border
            cell.alignment = data_align
            if r % 2 == 0:
                cell.fill = row_fill_even

    # Column widths
    col_widths = [12, 10, 16, 20, 14, 14, 10]
    for col_idx, width in enumerate(col_widths, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws1.row_dimensions[1].height = 22
    ws1.freeze_panes = 'A2'

    # ── Sheet 2: RegionData ────────────────────────────────────────────────
    ws2 = wb.create_sheet('RegionData')

    rdata_header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
    rdata_header_fill = PatternFill(start_color='FF375623', end_color='FF375623', fill_type='solid')

    region_headers = ['North', 'South', 'East', 'West']
    for col, rh in enumerate(region_headers, 1):
        cell = ws2.cell(row=1, column=col, value=rh)
        cell.font = rdata_header_font
        cell.fill = rdata_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = cell_border

    # Subregion lists per region (plain data — task will create named ranges)
    subregions = {
        'North': ['Minnesota', 'Wisconsin', 'Michigan', 'Illinois', 'Ohio'],
        'South': ['Florida', 'Georgia', 'Texas', 'Alabama', 'Louisiana'],
        'East':  ['New York', 'Massachusetts', 'Pennsylvania', 'New Jersey', 'Connecticut'],
        'West':  ['California', 'Washington', 'Oregon', 'Nevada', 'Arizona'],
    }

    for col_idx, region in enumerate(region_headers, 1):
        for row_idx, subregion in enumerate(subregions[region], 2):
            cell = ws2.cell(row=row_idx, column=col_idx, value=subregion)
            cell.border = cell_border
            cell.alignment = Alignment(horizontal='left', vertical='center')

    for col_idx in range(1, 5):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 16

    ws2.row_dimensions[1].height = 20

    # ── Save ───────────────────────────────────────────────────────────────
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
