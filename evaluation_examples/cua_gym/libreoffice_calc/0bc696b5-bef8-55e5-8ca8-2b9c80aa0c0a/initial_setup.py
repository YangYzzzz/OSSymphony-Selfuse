"""
Initial Setup: Apply data validation to budget spreadsheet
Task ID: calc_gcv_077
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_077'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget_Allocation"

    # Headers
    headers = ["Department", "Project", "Allocation", "Requested Budget"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 34 rows of realistic department/project data with Allocation values
    data = [
        ["Engineering", "Cloud Migration", 85000],
        ["Marketing", "Brand Refresh Campaign", 42000],
        ["Human Resources", "Employee Wellness Program", 18000],
        ["Finance", "ERP System Upgrade", 95000],
        ["Sales", "CRM Integration", 67000],
        ["IT Infrastructure", "Network Security Overhaul", 78000],
        ["Product Development", "Mobile App v3.0", 100000],
        ["Customer Support", "Helpdesk Automation", 32000],
        ["Legal", "Compliance Audit", 25000],
        ["Operations", "Warehouse Optimization", 54000],
        ["Research", "AI/ML Pilot Study", 88000],
        ["Quality Assurance", "Automated Testing Suite", 41000],
        ["Design", "UI/UX Redesign", 36000],
        ["Procurement", "Vendor Management Portal", 29000],
        ["Training", "Leadership Development", 15000],
        ["Facilities", "Office Renovation Phase 2", 72000],
        ["Data Analytics", "Business Intelligence Dashboard", 63000],
        ["Compliance", "GDPR Implementation", 47000],
        ["Public Relations", "Media Outreach Strategy", 22000],
        ["Supply Chain", "Logistics Tracking System", 58000],
        ["Accounting", "Financial Reporting Automation", 34000],
        ["Security", "Access Control Upgrade", 51000],
        ["Administration", "Document Management System", 19000],
        ["Business Development", "Partnership Expansion", 76000],
        ["Technical Writing", "Knowledge Base Overhaul", 12000],
        ["Internal Audit", "Risk Assessment Framework", 28000],
        ["Environmental Health", "Safety Compliance Training", 8000],
        ["Corporate Strategy", "Market Analysis Initiative", 91000],
        ["Investor Relations", "Annual Report Production", 5000],
        ["Customer Success", "Onboarding Workflow Redesign", 44000],
        ["DevOps", "CI/CD Pipeline Enhancement", 69000],
        ["Content Marketing", "SEO Content Strategy", 37000],
        ["Executive Office", "Board Meeting Technology", 53000],
        ["Field Operations", "Regional Expansion Support", 61000],
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])  # Department
        ws.cell(row=r, column=2, value=row_data[1])  # Project
        ws.cell(row=r, column=3, value=row_data[2])  # Allocation
        # Column D (Requested Budget) left EMPTY intentionally

    # Adjust column widths for readability
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 18

    # NO data validation - that is the task for the agent to add

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
