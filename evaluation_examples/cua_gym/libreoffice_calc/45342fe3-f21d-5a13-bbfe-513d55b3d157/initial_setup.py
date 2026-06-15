"""
Initial Setup: Create a spreadsheet with 'Data' sheet containing A1:E100 data,
some text fields containing commas (to test CSV export handling).
Task ID: calc_mcp_084
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_084'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Data'

    # Headers in row 1
    headers = ['Employee Name', 'Department', 'Job Title', 'Annual Salary', 'Start Date']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic data generation
    first_names = [
        'Sarah', 'Marcus', 'Elena', 'James', 'Priya', 'David', 'Mei',
        'Carlos', 'Fatima', 'Robert', 'Aisha', 'Thomas', 'Yuki', 'Daniel',
        'Olivia', 'Ahmed', 'Sophie', 'Brian', 'Nadia', 'Kevin',
        'Laura', 'Hassan', 'Emma', 'Victor', 'Zara'
    ]
    last_names = [
        'Chen', 'Johnson', 'Petrov', 'Williams', 'Sharma', 'Kim', 'Garcia',
        'Mueller', 'Okafor', 'Singh', 'Brown', 'Taylor', 'Yamamoto', 'Wilson',
        'Martinez', 'Li', 'Anderson', 'Davis', 'Thompson', 'Robinson',
        'Clark', 'Lewis', 'Walker', 'Hall', 'Young'
    ]

    departments = [
        'Engineering', 'Marketing', 'Sales', 'Human Resources',
        'Finance', 'Operations', 'Legal', 'Customer Support',
        'Research, Development', 'Product Management',
        'Quality Assurance, Testing', 'Data Science',
        'IT Infrastructure', 'Business Development',
        'Public Relations, Communications'
    ]

    # Job titles - some deliberately contain commas to test CSV handling
    job_titles = [
        'Software Engineer', 'Marketing Manager', 'Sales Representative',
        'HR Coordinator', 'Financial Analyst', 'Operations Director',
        'Legal Counsel', 'Support Specialist', 'Research Scientist',
        'Product Manager', 'QA Engineer', 'Data Analyst',
        'Systems Administrator', 'Business Analyst', 'PR Specialist',
        'Senior Developer, Backend', 'VP, Marketing',
        'Director, Sales Operations', 'Manager, Client Relations',
        'Lead Engineer, Infrastructure', 'Head of Design, UX',
        'Coordinator, Events and Outreach', 'Analyst, Risk and Compliance',
        'Specialist, Training and Development', 'Consultant, Strategy and Growth'
    ]

    random.seed(42)  # reproducible

    years = list(range(2018, 2026))
    months = list(range(1, 13))

    for row in range(2, 101):  # rows 2-100 = 99 data rows
        fn = first_names[(row - 2) % len(first_names)]
        ln = last_names[(row - 2) // len(first_names) % len(last_names)]
        name = f'{fn} {ln}'

        dept = departments[(row - 2) % len(departments)]
        title = job_titles[(row - 2) % len(job_titles)]
        salary = random.randint(45000, 185000)

        yr = years[random.randint(0, len(years) - 1)]
        mo = months[random.randint(0, len(months) - 1)]
        day = random.randint(1, 28)
        start_date = f'{yr}-{mo:02d}-{day:02d}'

        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=dept)
        ws.cell(row=row, column=3, value=title)
        ws.cell(row=row, column=4, value=salary)
        ws.cell(row=row, column=5, value=start_date)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 14

    # Make sure Documents directory does NOT have the CSV
    # (it should not exist yet, but be safe)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Ensure /home/user/Documents exists but no CSV file present
    os.makedirs(f'{WORKDIR}/Documents', exist_ok=True)
    csv_path = f'{WORKDIR}/Documents/data_semicolon.csv'
    if os.path.exists(csv_path):
        os.remove(csv_path)

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
