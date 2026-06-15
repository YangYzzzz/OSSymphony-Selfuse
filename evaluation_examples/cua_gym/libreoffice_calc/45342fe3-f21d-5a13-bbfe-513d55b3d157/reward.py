"""
Reward Script: Export 'Data' sheet as CSV with semicolon delimiters and UTF-8 encoding
Task ID: calc_mcp_084
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): CSV exists at correct path, uses semicolons, correct header
  Component 2 (0.35): Correct row count and data values match source xlsx
  Component 3 (0.35): UTF-8 encoding and comma-containing fields handled properly
"""

import os

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_084'
CSV_PATH = '/home/user/Documents/data_semicolon.csv'
XLSX_PATH = '/home/user/calc_mcp_084.xlsx'

EXPECTED_HEADER_FIELDS = ['Employee Name', 'Department', 'Job Title', 'Annual Salary', 'Start Date']


def verify_task():
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Early exit: CSV file must exist (precondition gate, not scored alone)
    if not os.path.exists(CSV_PATH):
        print(f"CRITICAL: CSV file not found at {CSV_PATH}")
        print("REWARD: 0.0")
        return 0.0

    # Read the CSV content (raw bytes and text)
    try:
        with open(CSV_PATH, 'rb') as f:
            raw_bytes = f.read()
        csv_text = raw_bytes.decode('utf-8')
        csv_lines = csv_text.strip().split('\n')
    except UnicodeDecodeError as e:
        print(f"CRITICAL: File is not valid UTF-8: {e}")
        print("REWARD: 0.0")
        return 0.0
    except Exception as e:
        print(f"CRITICAL: Cannot read CSV file: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Load the source xlsx for comparison
    try:
        import openpyxl
        wb = openpyxl.load_workbook(XLSX_PATH)
        ws = wb['Data']
        xlsx_max_row = ws.max_row
        xlsx_max_col = ws.max_column
    except Exception as e:
        print(f"CRITICAL: Cannot load source xlsx: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: CSV uses semicolon delimiters and has correct header (0.30 points)
    try:
        header_line = csv_lines[0]
        header_fields = header_line.split(';')
        header_match = [f.strip() for f in header_fields] == EXPECTED_HEADER_FIELDS

        # Check that semicolons are used as delimiters (header should have 4 semicolons for 5 fields)
        semicolon_count = header_line.count(';')
        uses_semicolons = semicolon_count == 4

        # Also verify that the first few data lines use semicolons consistently
        data_lines_ok = True
        for line in csv_lines[1:min(6, len(csv_lines))]:
            if line.count(';') < 4:
                data_lines_ok = False
                break

        if header_match and uses_semicolons and data_lines_ok:
            print(f"PASS: Component 1 -- Semicolon delimiters, correct header ({header_fields}) (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 -- header_match={header_match}, uses_semicolons={uses_semicolons}, data_lines_ok={data_lines_ok}")
            if not header_match:
                print(f"  Expected header: {EXPECTED_HEADER_FIELDS}")
                print(f"  Got header: {header_fields}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Correct row count and data values match source xlsx (0.35 points)
    try:
        # Expected: 1 header + (max_row - 1) data rows = max_row lines total
        expected_line_count = xlsx_max_row  # row 1 is header in xlsx, so total data rows = max_row - 1, plus 1 header line
        actual_line_count = len(csv_lines)

        row_count_ok = actual_line_count == expected_line_count

        # Verify a sample of data values against the xlsx
        values_correct = 0
        values_checked = 0
        sample_rows = [2, 10, 25, 50, 75, 99]  # check these xlsx rows (1-indexed)
        for xlsx_row in sample_rows:
            if xlsx_row > xlsx_max_row:
                continue
            csv_line_idx = xlsx_row - 1  # csv line index (0-indexed, line 0 is header)
            if csv_line_idx >= len(csv_lines):
                continue
            values_checked += 1

            # Get xlsx values for this row
            xlsx_vals = []
            for c in range(1, xlsx_max_col + 1):
                v = ws.cell(row=xlsx_row, column=c).value
                if v is None:
                    xlsx_vals.append('')
                else:
                    xlsx_vals.append(str(v))

            # Parse CSV line by semicolons
            csv_vals = csv_lines[csv_line_idx].split(';')
            csv_vals = [v.strip() for v in csv_vals]

            # Compare (strip whitespace, handle quoting)
            match = True
            for xv, cv in zip(xlsx_vals, csv_vals):
                xv_clean = xv.strip().strip('"')
                cv_clean = cv.strip().strip('"')
                if xv_clean != cv_clean:
                    match = False
                    break
            if match:
                values_correct += 1

        values_ratio = values_correct / max(values_checked, 1)

        if row_count_ok and values_ratio >= 0.8:
            print(f"PASS: Component 2 -- {actual_line_count} lines (expected {expected_line_count}), {values_correct}/{values_checked} sample rows match (0.35 pts)")
            total_score += 0.35
        elif row_count_ok or values_ratio >= 0.5:
            partial = 0.15
            print(f"PARTIAL: Component 2 -- row_count_ok={row_count_ok} ({actual_line_count}/{expected_line_count}), values={values_correct}/{values_checked} ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- {actual_line_count} lines (expected {expected_line_count}), {values_correct}/{values_checked} sample rows match")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: UTF-8 encoding and comma-containing fields handled properly (0.35 points)
    try:
        # Check 1: File is valid UTF-8 (already confirmed above by successful decode)
        utf8_ok = True

        # Check 2: Fields containing commas are preserved correctly
        # Find rows in xlsx that have commas in text, verify they appear correctly in CSV
        comma_fields_correct = 0
        comma_fields_checked = 0

        for xlsx_row in range(2, xlsx_max_row + 1):
            for c in range(1, xlsx_max_col + 1):
                v = ws.cell(row=xlsx_row, column=c).value
                if isinstance(v, str) and ',' in v:
                    comma_fields_checked += 1
                    if comma_fields_checked > 15:
                        break  # check up to 15 comma fields
                    # Check if this value appears in the corresponding CSV line
                    csv_line_idx = xlsx_row - 1
                    if csv_line_idx < len(csv_lines):
                        # The value should appear in the CSV line (possibly quoted)
                        csv_line = csv_lines[csv_line_idx]
                        # Strip possible quotes around the value
                        if v in csv_line or f'"{v}"' in csv_line:
                            comma_fields_correct += 1
            if comma_fields_checked > 15:
                break

        comma_ratio = comma_fields_correct / max(comma_fields_checked, 1)

        if utf8_ok and comma_ratio >= 0.8:
            print(f"PASS: Component 3 -- UTF-8 valid, {comma_fields_correct}/{comma_fields_checked} comma-containing fields correct (0.35 pts)")
            total_score += 0.35
        elif utf8_ok and comma_ratio >= 0.5:
            partial = 0.15
            print(f"PARTIAL: Component 3 -- UTF-8 valid, {comma_fields_correct}/{comma_fields_checked} comma-containing fields correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 -- utf8_ok={utf8_ok}, comma_fields={comma_fields_correct}/{comma_fields_checked}")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
verify_task()
