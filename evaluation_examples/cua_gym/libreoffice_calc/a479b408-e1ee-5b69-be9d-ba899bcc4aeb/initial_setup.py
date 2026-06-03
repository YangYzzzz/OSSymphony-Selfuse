"""
Initial Setup: Create spreadsheet with mutual fund performance data (no chart)
Task ID: calc_chart_line_multiple_series_054
Domain: libreoffice_calc
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_line_multiple_series_054'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: FundPerformance ---
    ws = wb.active
    ws.title = 'FundPerformance'

    # Headers (Row 1)
    headers = ['Quarter', 'Growth Fund', 'Income Fund', 'Balanced Fund', 'Index Fund']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Data rows (Rows 2-9) — exact values from task context
    data = [
        ['Q1 2023', 100,   100,   100,   100],
        ['Q2 2023', 103.2, 101.8, 102.4, 102.9],
        ['Q3 2023', 98.4,  103.1, 101.2, 100.1],
        ['Q4 2023', 108.6, 104.5, 106.8, 107.4],
        ['Q1 2024', 112.1, 105.8, 108.9, 110.2],
        ['Q2 2024', 118.4, 107.2, 112.3, 115.8],
        ['Q3 2024', 115.2, 108.6, 111.0, 112.4],
        ['Q4 2024', 124.8, 109.9, 116.5, 121.3],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # No charts in initial file — agent must create one

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheet: FundPerformance')
    print('Rows: 9 (1 header + 8 data rows)')
    print('Columns: Quarter, Growth Fund, Income Fund, Balanced Fund, Index Fund')
    print('Charts: None (task requires agent to create chart)')


create_initial()
