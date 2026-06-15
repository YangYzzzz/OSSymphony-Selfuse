"""
Reward Script: Heat map via conditional formatting on a pivot table
Task ID: calc_gcp_053
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): PivotTable sheet exists with correct dimensions
  Component 2 (0.25): Correct row/column labels (7 locations x 5 service types)
  Component 3 (0.25): Data cells contain numeric averages in valid range
  Component 4 (0.25): 3-color scale conditional formatting on data range
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_053'

EXPECTED_LOCATIONS = {'Downtown', 'Suburb-N', 'Suburb-S', 'Mall-East', 'Mall-West', 'Airport', 'Online'}
EXPECTED_SERVICES = {'In-Store', 'Phone', 'Email', 'Chat', 'Self-Service'}


def find_pivot_sheet(wb):
    """Find a sheet that looks like a pivot table (not the raw data sheet)."""
    for name in wb.sheetnames:
        if name == 'CustSatisfaction':
            continue
        ws = wb[name]
        # Check if it has a small number of rows (pivot-like) with location/service data
        if ws.max_row <= 20:
            return ws
    return None


def find_header_row(ws):
    """Find the row containing service type headers."""
    for r in range(1, ws.max_row + 1):
        row_vals = set()
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v and isinstance(v, str):
                row_vals.add(v.strip())
        # Check if this row contains at least 4 of the 5 expected service types
        if len(row_vals & EXPECTED_SERVICES) >= 4:
            return r
    return None


def find_data_range(ws, header_row):
    """Given header row, find the data start row, location col, and service cols."""
    # Service type columns
    service_cols = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v and isinstance(v, str) and v.strip() in EXPECTED_SERVICES:
            service_cols[v.strip()] = c

    # Location column: the column in header_row that has a label like "Store Location" or similar
    loc_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v and isinstance(v, str) and v.strip() not in EXPECTED_SERVICES:
            loc_col = c
            break

    # Data rows start right after header row
    data_start = header_row + 1
    # Find how many rows have location labels
    locations_found = {}
    for r in range(data_start, ws.max_row + 1):
        v = ws.cell(row=r, column=loc_col).value if loc_col else None
        if v and isinstance(v, str) and v.strip() in EXPECTED_LOCATIONS:
            locations_found[v.strip()] = r

    return loc_col, service_cols, locations_found


def verify_task(file_path):
    """Verify task completion with progressive scoring. Returns float 0.0-1.0."""
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: PivotTable sheet exists with correct structure (0.25 points)
    try:
        pivot_ws = find_pivot_sheet(wb)
        if pivot_ws is None:
            print("FAIL: Component 1 — No pivot table sheet found (only CustSatisfaction)")
        else:
            header_row = find_header_row(pivot_ws)
            if header_row is not None:
                print(f"PASS: Component 1 — Pivot sheet '{pivot_ws.title}' exists with headers at row {header_row} (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 1 — Pivot sheet '{pivot_ws.title}' found but no service type headers")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Early exit if no pivot sheet or headers
    if total_score == 0.0:
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Component 2: Correct row/column labels (0.25 points)
    try:
        loc_col, service_cols, locations_found = find_data_range(pivot_ws, header_row)
        services_ok = len(set(service_cols.keys()) & EXPECTED_SERVICES) >= 4
        locations_ok = len(set(locations_found.keys()) & EXPECTED_LOCATIONS) >= 6

        if services_ok and locations_ok:
            print(f"PASS: Component 2 — {len(service_cols)} service types, {len(locations_found)} locations (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — Services: {list(service_cols.keys())}, Locations: {list(locations_found.keys())}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Data cells contain numeric averages in valid range 1.0-5.0 (0.25 points)
    try:
        numeric_count = 0
        valid_range_count = 0
        total_data_cells = len(locations_found) * len(service_cols)

        for loc_name, row_num in locations_found.items():
            for svc_name, col_num in service_cols.items():
                val = pivot_ws.cell(row=row_num, column=col_num).value
                if val is not None and isinstance(val, (int, float)):
                    numeric_count += 1
                    if 1.0 <= float(val) <= 5.0:
                        valid_range_count += 1

        if total_data_cells > 0 and numeric_count >= total_data_cells * 0.8 and valid_range_count >= numeric_count * 0.8:
            print(f"PASS: Component 3 — {numeric_count}/{total_data_cells} numeric cells, {valid_range_count} in range 1-5 (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 3 — {numeric_count}/{total_data_cells} numeric, {valid_range_count} in valid range")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: 3-color scale conditional formatting on data cells (0.25 points)
    try:
        cf_rules = list(pivot_ws.conditional_formatting)
        color_scale_range = None

        for cf in cf_rules:
            for rule in cf.rules:
                if rule.type == 'colorScale' and rule.colorScale:
                    cs = rule.colorScale
                    # Must be a 3-color scale (3 cfvo entries and 3 colors)
                    if len(cs.cfvo) == 3 and len(cs.color) == 3:
                        color_scale_range = str(cf)
                        break
            if color_scale_range is not None:
                break

        if color_scale_range is not None:
            print(f"PASS: Component 4 — 3-color scale found on range {color_scale_range} (0.25 pts)")
            total_score += 0.25
        elif len(cf_rules) == 0:
            # Check if there are any CF rules at all
            if len(cf_rules) == 0:
                print("FAIL: Component 4 — No conditional formatting rules found")
            else:
                rule_types = []
                for cf in cf_rules:
                    for rule in cf.rules:
                        rule_types.append(rule.type)
                print(f"FAIL: Component 4 — CF rules found ({rule_types}) but no 3-color scale")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
