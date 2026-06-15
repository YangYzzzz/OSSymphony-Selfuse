"""
Initial Setup: Create CustSatisfaction data for heat map pivot table task
Task ID: calc_gcp_053
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_053'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'CustSatisfaction'

    # Headers
    headers = ['ResponseID', 'Date', 'StoreLocation', 'ServiceType', 'SatisfactionScore', 'WaitTime']
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data parameters
    store_locations = ['Downtown', 'Suburb-N', 'Suburb-S', 'Mall-East', 'Mall-West', 'Airport', 'Online']
    service_types = ['In-Store', 'Phone', 'Email', 'Chat', 'Self-Service']

    # Base satisfaction scores per store-service combination (for realistic patterns)
    base_scores = {
        'Downtown': {'In-Store': 3.8, 'Phone': 3.2, 'Email': 3.5, 'Chat': 3.9, 'Self-Service': 3.6},
        'Suburb-N': {'In-Store': 4.1, 'Phone': 3.5, 'Email': 3.3, 'Chat': 3.7, 'Self-Service': 3.8},
        'Suburb-S': {'In-Store': 3.5, 'Phone': 2.9, 'Email': 3.1, 'Chat': 3.4, 'Self-Service': 3.2},
        'Mall-East': {'In-Store': 3.3, 'Phone': 3.0, 'Email': 2.8, 'Chat': 3.6, 'Self-Service': 3.1},
        'Mall-West': {'In-Store': 4.2, 'Phone': 3.8, 'Email': 3.6, 'Chat': 4.0, 'Self-Service': 3.9},
        'Airport': {'In-Store': 2.5, 'Phone': 2.2, 'Email': 2.0, 'Chat': 2.8, 'Self-Service': 2.3},
        'Online': {'In-Store': 3.0, 'Phone': 3.4, 'Email': 3.7, 'Chat': 4.3, 'Self-Service': 4.1},
    }

    # Wait time base per service type (minutes)
    wait_base = {'In-Store': 12, 'Phone': 8, 'Email': 24, 'Chat': 3, 'Self-Service': 1}

    start_date = datetime(2024, 1, 1)
    end_date = datetime(2025, 3, 31)
    date_range_days = (end_date - start_date).days

    for i in range(1, 601):
        row = i + 1
        store = random.choice(store_locations)
        service = random.choice(service_types)

        # ResponseID
        ws.cell(row=row, column=1, value=i)

        # Date
        rand_date = start_date + timedelta(days=random.randint(0, date_range_days))
        ws.cell(row=row, column=2, value=rand_date.strftime('%Y-%m-%d'))

        # StoreLocation
        ws.cell(row=row, column=3, value=store)

        # ServiceType
        ws.cell(row=row, column=4, value=service)

        # SatisfactionScore (based on base + noise, clamped 1.0-5.0)
        base = base_scores[store][service]
        score = round(max(1.0, min(5.0, base + random.gauss(0, 0.6))), 1)
        ws.cell(row=row, column=5, value=score)

        # WaitTime (minutes, based on service type + noise)
        wt = max(0, round(wait_base[service] + random.gauss(0, wait_base[service] * 0.3)))
        ws.cell(row=row, column=6, value=wt)

    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 12

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
