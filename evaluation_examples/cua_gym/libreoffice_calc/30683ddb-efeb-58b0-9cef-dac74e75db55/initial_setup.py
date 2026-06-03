"""
Initial Setup: Cash flow projection spreadsheet with income/expense data
Task ID: calc_wf_041
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_041'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Flow"

    # --- Header styling ---
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # --- Row 1: Headers ---
    headers = ["Category"] + [f"Week {i}" for i in range(1, 13)]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Row layout ---
    # Row 2: Beginning Balance
    # Row 3: Sales (income)
    # Row 4: Services (income)
    # Row 5: Other Income
    # Row 6: Total Income (EMPTY - task asks agent to calculate)
    # Row 7: Payroll (expense)
    # Row 8: Rent (expense)
    # Row 9: Supplies (expense)
    # Row 10: Utilities (expense)
    # Row 11: Other Expenses
    # Row 12: Total Expenses (EMPTY - task asks agent to calculate)
    # Row 13: Net Cash Flow (EMPTY - task asks agent to calculate)
    # Row 14: Running Balance (EMPTY - task asks agent to calculate)

    row_labels = [
        "Beginning Balance",
        "Sales",
        "Services",
        "Other Income",
        "Total Income",
        "Payroll",
        "Rent",
        "Supplies",
        "Utilities",
        "Other Expenses",
        "Total Expenses",
        "Net Cash Flow",
        "Running Balance",
    ]

    label_font = Font(name="Calibri", size=11, bold=True)
    section_fills = {
        "Beginning Balance": PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid"),
        "Total Income": PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid"),
        "Total Expenses": PatternFill(start_color="FFFCE4D6", end_color="FFFCE4D6", fill_type="solid"),
        "Net Cash Flow": PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid"),
        "Running Balance": PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid"),
    }

    for r, label in enumerate(row_labels, 2):
        cell = ws.cell(row=r, column=1, value=label)
        cell.font = label_font
        cell.border = thin_border
        if label in section_fills:
            cell.fill = section_fills[label]
            # Apply fill across entire row for section rows
            for col in range(2, 14):
                ws.cell(row=r, column=col).fill = section_fills[label]
                ws.cell(row=r, column=col).border = thin_border

    # --- Beginning Balance: $25,000 in Week 1 ---
    ws.cell(row=2, column=2, value=25000)
    ws.cell(row=2, column=2).number_format = '$#,##0.00'
    ws.cell(row=2, column=2).border = thin_border

    # --- Weekly income data (realistic, varied) ---
    # Sales (Row 3)
    sales_data = [8500, 9200, 7800, 10500, 8900, 11200, 9600, 8100, 10800, 9400, 7500, 11500]
    # Services (Row 4)
    services_data = [3200, 2800, 3500, 3100, 4200, 3800, 2900, 3600, 3300, 4100, 3700, 3000]
    # Other Income (Row 5)
    other_income_data = [500, 200, 800, 150, 600, 300, 450, 100, 750, 350, 200, 900]

    income_rows = {3: sales_data, 4: services_data, 5: other_income_data}
    for row_num, data in income_rows.items():
        for col, val in enumerate(data, 2):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.number_format = '$#,##0.00'
            cell.border = thin_border

    # --- Weekly expense data (realistic, varied) ---
    # Payroll (Row 7) - largest expense, fairly stable
    payroll_data = [6500, 6500, 6500, 6500, 6500, 6500, 6500, 6500, 6500, 6500, 6500, 6500]
    # Rent (Row 8) - fixed
    rent_data = [3500, 3500, 3500, 3500, 3500, 3500, 3500, 3500, 3500, 3500, 3500, 3500]
    # Supplies (Row 9)
    supplies_data = [1200, 800, 1500, 950, 1100, 1800, 900, 1350, 1050, 1400, 1600, 1100]
    # Utilities (Row 10)
    utilities_data = [850, 920, 780, 1050, 890, 950, 1100, 870, 930, 1020, 860, 980]
    # Other Expenses (Row 11)
    other_exp_data = [400, 650, 300, 1200, 500, 350, 800, 450, 600, 300, 750, 400]

    expense_rows = {7: payroll_data, 8: rent_data, 9: supplies_data, 10: utilities_data, 11: other_exp_data}
    for row_num, data in expense_rows.items():
        for col, val in enumerate(data, 2):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.number_format = '$#,##0.00'
            cell.border = thin_border

    # Row 6 (Total Income), Row 12 (Total Expenses), Row 13 (Net Cash Flow), Row 14 (Running Balance)
    # are left EMPTY - the task instructs the agent to add formulas here

    # --- Column widths ---
    ws.column_dimensions["A"].width = 20
    for col_letter in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]:
        ws.column_dimensions[col_letter].width = 14

    # --- Freeze first column and header row ---
    ws.freeze_panes = "B2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
