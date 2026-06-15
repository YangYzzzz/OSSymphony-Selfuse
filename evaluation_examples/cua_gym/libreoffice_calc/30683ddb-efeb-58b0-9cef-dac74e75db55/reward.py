"""
Reward Script: Cash Flow Projection Spreadsheet
Task ID: calc_wf_041
Domain: libreoffice_calc
Scoring:
  Component 1: Total Income formulas (Row 6)     — 0.20 pts
  Component 2: Total Expenses formulas (Row 12)   — 0.20 pts
  Component 3: Net Cash Flow formulas (Row 13)    — 0.15 pts
  Component 4: Running Balance formulas (Row 14)  — 0.15 pts
  Component 5: Conditional formatting on balance  — 0.15 pts
  Component 6: Area chart + threshold line        — 0.15 pts
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_041'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Cash Flow' sheet must exist
    if 'Cash Flow' not in wb.sheetnames:
        print("FAIL: 'Cash Flow' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Cash Flow']

    # Helper to normalize formula strings for comparison
    def norm(f):
        if not isinstance(f, str):
            return ''
        return f.upper().replace(' ', '')

    # ---------------------------------------------------------------
    # Component 1: Total Income formulas in Row 6 (0.20 points)
    # Row 6 should have =SUM(Xn3:Xn5) for each week column B-M
    # Initial env has these as None — only golden has formulas
    # ---------------------------------------------------------------
    try:
        income_pass = 0
        for col_idx in range(2, 14):  # columns B(2) through M(13)
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            expected = f'=SUM({col_letter}3:{col_letter}5)'
            actual = ws.cell(row=6, column=col_idx).value
            if norm(actual) == norm(expected):
                income_pass += 1
        if income_pass == 12:
            print(f"PASS: Component 1 — All 12 Total Income SUM formulas correct (0.20 pts)")
            total_score += 0.20
        elif income_pass >= 6:
            partial = round(0.20 * income_pass / 12, 2)
            print(f"PARTIAL: Component 1 — {income_pass}/12 Total Income formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {income_pass}/12 Total Income formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ---------------------------------------------------------------
    # Component 2: Total Expenses formulas in Row 12 (0.20 points)
    # Row 12 should have =SUM(Xn7:Xn11) for each week column B-M
    # Initial env has these as None — only golden has formulas
    # ---------------------------------------------------------------
    try:
        expense_pass = 0
        for col_idx in range(2, 14):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            expected = f'=SUM({col_letter}7:{col_letter}11)'
            actual = ws.cell(row=12, column=col_idx).value
            if norm(actual) == norm(expected):
                expense_pass += 1
        if expense_pass == 12:
            print(f"PASS: Component 2 — All 12 Total Expenses SUM formulas correct (0.20 pts)")
            total_score += 0.20
        elif expense_pass >= 6:
            partial = round(0.20 * expense_pass / 12, 2)
            print(f"PARTIAL: Component 2 — {expense_pass}/12 Total Expenses formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {expense_pass}/12 Total Expenses formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---------------------------------------------------------------
    # Component 3: Net Cash Flow formulas in Row 13 (0.15 points)
    # Row 13 should have =Xn6-Xn12 for each week column B-M
    # Initial env has these as None — only golden has formulas
    # ---------------------------------------------------------------
    try:
        net_pass = 0
        for col_idx in range(2, 14):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            expected = f'={col_letter}6-{col_letter}12'
            actual = ws.cell(row=13, column=col_idx).value
            if norm(actual) == norm(expected):
                net_pass += 1
        if net_pass == 12:
            print(f"PASS: Component 3 — All 12 Net Cash Flow formulas correct (0.15 pts)")
            total_score += 0.15
        elif net_pass >= 6:
            partial = round(0.15 * net_pass / 12, 2)
            print(f"PARTIAL: Component 3 — {net_pass}/12 Net Cash Flow formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {net_pass}/12 Net Cash Flow formulas found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ---------------------------------------------------------------
    # Component 4: Running Balance formulas in Row 14 (0.15 points)
    # B14 = =B2+B13 (beginning balance + week 1 net flow)
    # C14 = =B14+C13, D14 = =C14+D13, ... M14 = =L14+M13
    # Initial env has these as None — only golden has formulas
    # ---------------------------------------------------------------
    try:
        balance_pass = 0
        # Check B14 specifically: =B2+B13
        actual_b14 = ws.cell(row=14, column=2).value
        if norm(actual_b14) == norm('=B2+B13'):
            balance_pass += 1

        # Check C14 through M14: =<prev_col>14+<col>13
        for col_idx in range(3, 14):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            prev_letter = openpyxl.utils.get_column_letter(col_idx - 1)
            expected = f'={prev_letter}14+{col_letter}13'
            actual = ws.cell(row=14, column=col_idx).value
            if norm(actual) == norm(expected):
                balance_pass += 1

        if balance_pass == 12:
            print(f"PASS: Component 4 — All 12 Running Balance formulas correct (0.15 pts)")
            total_score += 0.15
        elif balance_pass >= 6:
            partial = round(0.15 * balance_pass / 12, 2)
            print(f"PARTIAL: Component 4 — {balance_pass}/12 Running Balance formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {balance_pass}/12 Running Balance formulas found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ---------------------------------------------------------------
    # Component 5: Conditional formatting on balance row (0.15 points)
    # B14:M14 should have cellIs lessThan 5000 with red fill
    # Initial env has NO conditional formatting
    # ---------------------------------------------------------------
    try:
        cf_found = False
        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            # Check if the conditional formatting applies to the balance row (row 14)
            if '14' in cf_range:
                for rule in cf.rules:
                    # Check for lessThan operator with value 5000
                    if rule.type == 'cellIs' and rule.operator == 'lessThan':
                        formula_str = str(rule.formula)
                        if '5000' in formula_str:
                            # Check for red-ish fill
                            has_red_fill = False
                            if rule.dxf and rule.dxf.fill:
                                try:
                                    rgb = rule.dxf.fill.fgColor.rgb
                                    if rgb and 'FF0000' in str(rgb).upper():
                                        has_red_fill = True
                                except Exception:
                                    pass
                            if has_red_fill:
                                cf_found = True
                                print(f"PASS: Component 5 — Conditional formatting on row 14: cellIs < 5000 with red fill (0.15 pts)")
                                total_score += 0.15
                            else:
                                # Partial: rule exists but fill may differ
                                cf_found = True
                                print(f"PARTIAL: Component 5 — Conditional formatting rule found but fill color not red (0.07 pts)")
                                total_score += 0.07
        if not cf_found:
            print(f"FAIL: Component 5 — No conditional formatting with cellIs < 5000 on balance row")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # ---------------------------------------------------------------
    # Component 6: Area chart + threshold line (0.15 points)
    # Should have at least 1 chart; primary = AreaChart with
    # a combined LineChart for the threshold. Initial has no charts.
    # ---------------------------------------------------------------
    try:
        charts = ws._charts
        if len(charts) == 0:
            print(f"FAIL: Component 6 — No charts found in 'Cash Flow' sheet")
        else:
            chart = charts[0]
            sub_charts = chart._charts
            has_area = False
            has_line = False
            for sc in sub_charts:
                sc_type = type(sc).__name__
                if 'Area' in sc_type:
                    has_area = True
                if 'Line' in sc_type:
                    has_line = True

            if has_area and has_line:
                print(f"PASS: Component 6 — Area chart with threshold line found (0.15 pts)")
                total_score += 0.15
            elif has_area:
                print(f"PARTIAL: Component 6 — Area chart found but no threshold line (0.07 pts)")
                total_score += 0.07
            elif len(charts) > 0:
                print(f"PARTIAL: Component 6 — Chart found but not an AreaChart ({type(chart).__name__}) (0.05 pts)")
                total_score += 0.05
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
