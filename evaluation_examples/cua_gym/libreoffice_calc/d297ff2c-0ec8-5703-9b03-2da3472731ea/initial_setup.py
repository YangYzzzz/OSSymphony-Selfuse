"""
Initial Setup: Create a split commission calculator for deals involving multiple reps
Task ID: calc_sales_093
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_093'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Splits ---
    ws = wb.active
    ws.title = 'Splits'

    # Headers
    headers = {
        'A1': 'Deal',
        'B1': 'Total Commission',
        'C1': 'Rep 1',
        'D1': 'Rep 1 Split %',
        'E1': 'Rep 1 Amount',
        'F1': 'Rep 2',
        'G1': 'Rep 2 Split %',
        'H1': 'Rep 2 Amount',
    }
    for coord, val in headers.items():
        ws[coord] = val

    # Summary headers
    ws['J1'] = 'Rep'
    ws['K1'] = 'Total Earned'

    # Style headers
    header_font = Font(bold=True)
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'J', 'K']:
        ws[f'{col_letter}1'].font = header_font

    # Data rows
    # Row 2: Deal D1
    ws['A2'] = 'D1'
    ws['B2'] = 15000
    ws['C2'] = 'Alice'
    ws['D2'] = 0.60
    # E2 intentionally empty (task: enter formula E=B*D)
    ws['F2'] = 'Bob'
    ws['G2'] = 0.40
    # H2 intentionally empty (task: enter formula H=B*G)

    # Row 3: Deal D2
    ws['A3'] = 'D2'
    ws['B3'] = 8000
    ws['C3'] = 'Carol'
    ws['D3'] = 0.50
    ws['F3'] = 'Dan'
    ws['G3'] = 0.50

    # Row 4: Deal D3
    ws['A4'] = 'D3'
    ws['B4'] = 22000
    ws['C4'] = 'Alice'
    ws['D4'] = 0.70
    ws['F4'] = 'Eve'
    ws['G4'] = 0.30

    # Row 5: Deal D4
    ws['A5'] = 'D4'
    ws['B5'] = 12000
    ws['C5'] = 'Bob'
    ws['D5'] = 0.55
    ws['F5'] = 'Carol'
    ws['G5'] = 0.45

    # Summary rep names (K column intentionally empty - task requires SUMIF formulas)
    ws['J2'] = 'Alice'
    ws['J3'] = 'Bob'
    ws['J4'] = 'Carol'
    ws['J5'] = 'Dan'
    ws['J6'] = 'Eve'

    # Format percentage columns
    for row in range(2, 6):
        ws.cell(row=row, column=4).number_format = '0%'
        ws.cell(row=row, column=7).number_format = '0%'

    # Format currency columns
    for row in range(2, 6):
        ws.cell(row=row, column=2).number_format = '$#,##0'

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['J'].width = 14
    ws.column_dimensions['K'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
