"""
Reward Script: Split commission calculator with per-rep amount formulas and summary totals.
Task ID: calc_sales_093
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Rep 1 Amount formulas (E2:E5) = B*D
  Component 2 (0.30): Rep 2 Amount formulas (H2:H5) = B*G
  Component 3 (0.40): Summary Total Earned (K2:K6) uses SUMIF across both rep columns
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_093'


def normalize_formula(f):
    """Normalize a formula string for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def is_multiplication_formula(formula, row, col_b, col_factor):
    """
    Check if formula multiplies col_b by col_factor for the given row.
    E.g., =B2*D2, =D2*B2, =$B$2*$D$2, etc.
    """
    norm = normalize_formula(formula)
    if not norm.startswith('='):
        return False

    # Remove the leading '='
    expr = norm[1:]

    # Build expected patterns (with optional $ signs)
    def cell_pat(col, r):
        return r'\$?' + col + r'\$?' + str(r)

    pat1 = cell_pat(col_b, row) + r'\*' + cell_pat(col_factor, row)
    pat2 = cell_pat(col_factor, row) + r'\*' + cell_pat(col_b, row)

    return bool(re.match(pat1 + '$', expr)) or bool(re.match(pat2 + '$', expr))


def has_sumif_for_rep(formula):
    """
    Check if formula contains SUMIF (or SUMPRODUCT) referencing both rep columns.
    We accept any formula that uses SUMIF or SUMPRODUCT and references
    columns C/E (Rep 1) and F/H (Rep 2).
    """
    norm = normalize_formula(formula)
    if not norm.startswith('='):
        return False

    # Must contain SUMIF or SUMPRODUCT
    has_agg = 'SUMIF(' in norm or 'SUMPRODUCT(' in norm
    if not has_agg:
        return False

    # Should reference both Rep 1 columns (C and E) and Rep 2 columns (F and H)
    has_rep1 = ('C' in norm and 'E' in norm)
    has_rep2 = ('F' in norm and 'H' in norm)

    return has_rep1 and has_rep2


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Splits' sheet must exist
    if 'Splits' not in wb.sheetnames:
        print("CRITICAL: 'Splits' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Splits']

    # Component 1: Rep 1 Amount formulas E2:E5 (0.30 points)
    # Each correct formula = 0.075 pts
    try:
        comp1_score = 0.0
        for row in range(2, 6):
            cell_val = ws.cell(row=row, column=5).value  # column E
            if is_multiplication_formula(cell_val, row, 'B', 'D'):
                print(f"PASS: E{row} has correct formula: {cell_val}")
                comp1_score += 0.075
            else:
                print(f"FAIL: E{row} expected =B{row}*D{row}, found: {cell_val}")
        if comp1_score > 0:
            print(f"PASS: Component 1 -- Rep 1 Amount formulas ({comp1_score:.3f} pts)")
            total_score += comp1_score
        else:
            print(f"FAIL: Component 1 -- No Rep 1 Amount formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Rep 2 Amount formulas H2:H5 (0.30 points)
    # Each correct formula = 0.075 pts
    try:
        comp2_score = 0.0
        for row in range(2, 6):
            cell_val = ws.cell(row=row, column=8).value  # column H
            if is_multiplication_formula(cell_val, row, 'B', 'G'):
                print(f"PASS: H{row} has correct formula: {cell_val}")
                comp2_score += 0.075
            else:
                print(f"FAIL: H{row} expected =B{row}*G{row}, found: {cell_val}")
        if comp2_score > 0:
            print(f"PASS: Component 2 -- Rep 2 Amount formulas ({comp2_score:.3f} pts)")
            total_score += comp2_score
        else:
            print(f"FAIL: Component 2 -- No Rep 2 Amount formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Summary Total Earned K2:K6 (0.40 points)
    # Each correct SUMIF formula = 0.08 pts
    try:
        comp3_score = 0.0
        for row in range(2, 7):
            cell_val = ws.cell(row=row, column=11).value  # column K
            if has_sumif_for_rep(cell_val):
                print(f"PASS: K{row} has valid SUMIF formula: {cell_val}")
                comp3_score += 0.08
            else:
                print(f"FAIL: K{row} expected SUMIF formula across both rep columns, found: {cell_val}")
        if comp3_score > 0:
            print(f"PASS: Component 3 -- Summary SUMIF formulas ({comp3_score:.3f} pts)")
            total_score += comp3_score
        else:
            print(f"FAIL: Component 3 -- No Summary SUMIF formulas found")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
