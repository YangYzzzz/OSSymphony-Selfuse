"""
Initial Setup: Create a spreadsheet with monthly revenue data for H1.
Task ID: calc_sales_041
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_041'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Revenue ---
    ws = wb.active
    ws.title = 'Revenue'

    # Headers
    ws.cell(row=1, column=1, value='Month')
    ws.cell(row=1, column=2, value='Revenue')

    # Style headers
    header_font = Font(name='Calibri', size=11, bold=True)
    for col in range(1, 3):
        ws.cell(row=1, column=col).font = header_font

    # Data rows - monthly revenue for H1
    data = [
        ['Jan', 85000],
        ['Feb', 92000],
        ['Mar', 78000],
        ['Apr', 105000],
        ['May', 98000],
        ['Jun', 115000],
    ]
    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        ws.cell(row=r, column=2, value=row_data[1])

    # Format revenue column as currency
    for r in range(2, 8):
        ws.cell(row=r, column=2).number_format = '$#,##0'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15

    # NO charts - that is the task for the agent

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
