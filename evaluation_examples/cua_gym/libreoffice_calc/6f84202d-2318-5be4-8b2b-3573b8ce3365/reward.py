"""
Reward Script: AVERAGEIFS for average order processing time by priority and warehouse
Task ID: calc_ops_030
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.25): H2 contains a valid AVERAGEIFS formula for WH-A/High
  - Component 2 (0.25): H3 contains a valid AVERAGEIFS formula for WH-A/Low
  - Component 3 (0.25): H4 contains a valid AVERAGEIFS formula for WH-B/High
  - Component 4 (0.25): H5 contains a valid AVERAGEIFS formula for WH-B/Low
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_030'


def normalize_formula(f):
    """Normalize formula for comparison: uppercase, strip spaces, remove leading =."""
    if not isinstance(f, str):
        return ''
    f = f.strip()
    if f.startswith('='):
        f = f[1:]
    return f.upper().replace(' ', '').replace('$', '')


def is_averageifs_formula(value):
    """Check if cell value is an AVERAGEIFS formula."""
    if not isinstance(value, str):
        return False
    norm = normalize_formula(value)
    return norm.startswith('AVERAGEIFS(')


def check_averageifs_references_correct_ranges(formula_str):
    """
    Verify the AVERAGEIFS formula references the data range D2:D8 (avg_range),
    B2:B8 (criteria_range1 for warehouse), and C2:C8 (criteria_range2 for priority).
    Returns True if all three ranges are present.
    """
    norm = normalize_formula(formula_str)
    # Check for the processing time range (D column)
    has_d_range = bool(re.search(r'D\d+:D\d+', norm))
    # Check for warehouse criteria range (B column)
    has_b_range = bool(re.search(r'B\d+:B\d+', norm))
    # Check for priority criteria range (C column)
    has_c_range = bool(re.search(r'C\d+:C\d+', norm))
    return has_d_range and has_b_range and has_c_range


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet exists
    if 'Processing' not in wb.sheetnames:
        print("CRITICAL: 'Processing' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Processing']

    # Expected ground truth values (from context):
    # H2: AVERAGEIFS for WH-A, High -> expected result 2.50
    # H3: AVERAGEIFS for WH-A, Low  -> expected result 6.00
    # H4: AVERAGEIFS for WH-B, High -> expected result 4.00
    # H5: AVERAGEIFS for WH-B, Low  -> expected result 7.50

    cells_to_check = {
        'H2': ('WH-A', 'High', 2.50),
        'H3': ('WH-A', 'Low', 6.00),
        'H4': ('WH-B', 'High', 4.00),
        'H5': ('WH-B', 'Low', 7.50),
    }

    for cell_ref, (warehouse, priority, expected_val) in cells_to_check.items():
        component_name = f"{cell_ref} AVERAGEIFS for {warehouse}/{priority}"
        try:
            cell_value = ws[cell_ref].value

            if cell_value is None:
                print(f"FAIL: {component_name} -- cell is empty (None)")
                continue

            # Check if it's an AVERAGEIFS formula
            if not is_averageifs_formula(cell_value):
                # Could be a computed numeric value (if agent used the GUI and file was saved)
                # Check if the numeric value matches expected
                try:
                    numeric_val = float(cell_value)
                    if abs(numeric_val - expected_val) < 0.01:
                        print(f"PASS: {component_name} -- numeric value {numeric_val} matches expected {expected_val} (0.25 pts)")
                        total_score += 0.25
                    else:
                        print(f"FAIL: {component_name} -- value {cell_value} is not AVERAGEIFS formula and numeric {numeric_val} != {expected_val}")
                except (ValueError, TypeError):
                    print(f"FAIL: {component_name} -- value {repr(cell_value)} is neither AVERAGEIFS formula nor matching number")
                continue

            # It's an AVERAGEIFS formula -- check ranges are correct
            if check_averageifs_references_correct_ranges(cell_value):
                print(f"PASS: {component_name} -- formula: {cell_value} (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: {component_name} -- AVERAGEIFS formula present but ranges seem incorrect: {cell_value}")

        except Exception as e:
            print(f"ERROR: {component_name} -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
