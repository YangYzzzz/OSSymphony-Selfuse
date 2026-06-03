"""
Initial Setup: Create order processing spreadsheet with AVERAGEIFS summary table (empty results)
Task ID: calc_ops_030
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_030'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Processing'

    # --- Headers row 1 ---
    headers = ['Order', 'Warehouse', 'Priority', 'Processing Time (hrs)']
    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    # --- Data rows 2-8 ---
    data = [
        ['O-01', 'WH-A', 'High', 2],
        ['O-02', 'WH-B', 'Low', 8],
        ['O-03', 'WH-A', 'High', 3],
        ['O-04', 'WH-A', 'Low', 6],
        ['O-05', 'WH-B', 'High', 4],
        ['O-06', 'WH-B', 'Low', 7],
        ['O-07', 'WH-A', 'High', 2.5],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # --- Summary table headers at F1:H1 ---
    summary_headers = ['Warehouse', 'Priority', 'Avg Time']
    for col_offset, h in enumerate(summary_headers):
        cell = ws.cell(row=1, column=6 + col_offset, value=h)
        cell.font = header_font

    # --- Summary table criteria F2:G5 ---
    summary_data = [
        ['WH-A', 'High'],
        ['WH-A', 'Low'],
        ['WH-B', 'High'],
        ['WH-B', 'Low'],
    ]
    for r, row_data in enumerate(summary_data, 2):
        ws.cell(row=r, column=6, value=row_data[0])
        ws.cell(row=r, column=7, value=row_data[1])

    # H2:H5 intentionally left EMPTY - task is to enter AVERAGEIFS formulas

    # --- Column widths for readability ---
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
