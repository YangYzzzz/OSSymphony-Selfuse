"""
Initial Setup: Product Registry spreadsheet with product codes (some duplicated, some unique)
Task ID: calc_fmt_condfmt_unique_values_089
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_condfmt_unique_values_089'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Product Registry'

    # Header row
    ws['A1'] = 'Product Code'

    # Product codes: rows 2-50 (49 entries)
    # Mix of unique and duplicate values
    product_codes = [
        'PRD-10042',   # row 2  - unique
        'PRD-20187',   # row 3  - duplicate (appears 2x)
        'PRD-30054',   # row 4  - unique
        'PRD-20187',   # row 5  - duplicate
        'PRD-40093',   # row 6  - unique
        'PRD-50021',   # row 7  - duplicate (appears 3x)
        'PRD-60178',   # row 8  - unique
        'PRD-70045',   # row 9  - duplicate (appears 2x)
        'PRD-80231',   # row 10 - unique
        'PRD-50021',   # row 11 - duplicate
        'PRD-90067',   # row 12 - unique
        'PRD-10142',   # row 13 - unique
        'PRD-70045',   # row 14 - duplicate
        'PRD-20389',   # row 15 - unique
        'PRD-30254',   # row 16 - duplicate (appears 2x)
        'PRD-40193',   # row 17 - unique
        'PRD-50121',   # row 18 - unique
        'PRD-30254',   # row 19 - duplicate
        'PRD-60278',   # row 20 - unique
        'PRD-50021',   # row 21 - duplicate
        'PRD-70145',   # row 22 - unique
        'PRD-80331',   # row 23 - duplicate (appears 2x)
        'PRD-90167',   # row 24 - unique
        'PRD-10242',   # row 25 - unique
        'PRD-80331',   # row 26 - duplicate
        'PRD-20487',   # row 27 - unique
        'PRD-30354',   # row 28 - duplicate (appears 2x)
        'PRD-40293',   # row 29 - unique
        'PRD-50221',   # row 30 - unique
        'PRD-30354',   # row 31 - duplicate
        'PRD-60378',   # row 32 - unique
        'PRD-70245',   # row 33 - unique
        'PRD-80431',   # row 34 - unique
        'PRD-90267',   # row 35 - duplicate (appears 2x)
        'PRD-10342',   # row 36 - unique
        'PRD-20587',   # row 37 - unique
        'PRD-90267',   # row 38 - duplicate
        'PRD-30454',   # row 39 - unique
        'PRD-40393',   # row 40 - unique
        'PRD-50321',   # row 41 - duplicate (appears 2x)
        'PRD-60478',   # row 42 - unique
        'PRD-50321',   # row 43 - duplicate
        'PRD-70345',   # row 44 - unique
        'PRD-80531',   # row 45 - unique
        'PRD-90367',   # row 46 - unique
        'PRD-10442',   # row 47 - unique
        'PRD-20687',   # row 48 - unique
        'PRD-30554',   # row 49 - unique
        'PRD-40493',   # row 50 - unique
    ]

    for i, code in enumerate(product_codes, start=2):
        ws.cell(row=i, column=1, value=code)

    # No conditional formatting on initial file
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
