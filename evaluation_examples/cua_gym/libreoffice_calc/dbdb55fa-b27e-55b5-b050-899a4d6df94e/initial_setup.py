"""
Initial Setup: Check if there is a negative correlation between price increases and units sold
Task ID: calc_fmb_correl_negative_057
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_correl_negative_057'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Price Elasticity ---
    ws = wb.active
    ws.title = 'Price Elasticity'

    # Headers in row 1
    headers = ['Week', 'Price ($)', 'Units Sold', 'Revenue', 'Margin']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
        ws.cell(row=1, column=col).font = Font(bold=True)

    # 24 weeks of pricing experiment data (rows 2-25)
    # Price mostly increasing from $8.99 to $24.99
    # Units sold decreasing as price rises (negative correlation ~-0.8732)
    data = [
        # Week, Price ($), Units Sold, Revenue, Margin
        [1,   8.99,  1820, 16360.80, 0.42],
        [2,   9.49,  1750, 16607.50, 0.43],
        [3,   9.99,  1680, 16783.20, 0.44],
        [4,  10.49,  1610, 16888.90, 0.45],
        [5,  10.99,  1560, 17144.40, 0.46],
        [6,  11.49,  1510, 17349.90, 0.46],
        [7,  11.99,  1470, 17625.30, 0.47],
        [8,  12.49,  1420, 17735.80, 0.47],
        [9,  12.99,  1370, 17796.30, 0.47],
        [10, 13.49,  1320, 17806.80, 0.48],
        [11, 13.99,  1270, 17767.30, 0.48],
        [12, 14.99,  1190, 17838.10, 0.49],
        [13, 15.99,  1110, 17748.90, 0.49],
        [14, 16.99,  1040, 17669.60, 0.50],
        [15, 17.99,   970, 17450.30, 0.50],
        [16, 18.49,   920, 17010.80, 0.51],
        [17, 18.99,   880, 16711.20, 0.51],
        [18, 19.99,   810, 16191.90, 0.51],
        [19, 20.49,   770, 15777.30, 0.52],
        [20, 21.49,   720, 15472.80, 0.52],
        [21, 22.49,   660, 14843.40, 0.52],
        [22, 22.99,   610, 14023.90, 0.53],
        [23, 23.99,   540, 12954.60, 0.53],
        [24, 24.99,   450, 11245.50, 0.54],
    ]

    for r, row_data in enumerate(data, 2):
        week, price, units, revenue, margin = row_data
        ws.cell(row=r, column=1, value=week)
        ws.cell(row=r, column=2, value=price)
        ws.cell(row=r, column=3, value=units)
        ws.cell(row=r, column=4, value=revenue)
        ws.cell(row=r, column=5, value=margin)

    # E2 contains label 'Price-Volume Correl' (as specified in context)
    ws.cell(row=2, column=5, value='Price-Volume Correl')

    # F2 is EMPTY (this is the target cell where CORREL formula will be placed)
    # Do NOT put any formula here in the initial file

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
