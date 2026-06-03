"""
Reward Script: Check negative correlation between price increases and units sold
Task ID: calc_fmb_correl_negative_057
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4 pts): Cell F2 contains a CORREL formula (function name present)
  Component 2 (0.3 pts): CORREL formula references price column B2:B25
  Component 3 (0.3 pts): CORREL formula references units sold column C2:C25
  Total: 1.0

Only F2 changes between initial (None) and golden (=CORREL(B2:B25,C2:C25)).
All components fail on initial file (F2 is empty) and pass on golden file.
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmb_correl_negative_057'
SHEET_NAME = 'Price Elasticity'
TARGET_CELL = 'F2'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Sheets present: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]
    f2_value = ws[TARGET_CELL].value

    print(f"Cell F2 value: {repr(f2_value)}")

    # Component 1: F2 contains a CORREL formula (0.4 points)
    # FAILS on initial (F2=None), PASSES on golden (=CORREL(...))
    try:
        has_correl_formula = (
            f2_value is not None
            and isinstance(f2_value, str)
            and 'CORREL' in f2_value.upper()
        )
        if has_correl_formula:
            print(f"PASS: Component 1 — F2 contains a CORREL formula: {repr(f2_value)} (0.4 pts)")
            total_score += 0.4
        else:
            if f2_value is None:
                print(f"FAIL: Component 1 — F2 is empty; expected a CORREL formula")
            else:
                print(f"FAIL: Component 1 — F2={repr(f2_value)} does not contain CORREL function")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: CORREL formula references price column B2:B25 (0.3 points)
    # FAILS on initial (F2=None, no formula), PASSES on golden
    try:
        # Normalize formula for checking: remove spaces, uppercase
        formula_normalized = f2_value.upper().replace(" ", "") if f2_value else ""
        has_b_range = (
            f2_value is not None
            and isinstance(f2_value, str)
            and 'B2:B25' in formula_normalized
        )
        if has_b_range:
            print(f"PASS: Component 2 — CORREL formula references B2:B25 (price column) (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — Expected B2:B25 in formula, got: {repr(f2_value)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: CORREL formula references units sold column C2:C25 (0.3 points)
    # FAILS on initial (F2=None, no formula), PASSES on golden
    try:
        has_c_range = (
            f2_value is not None
            and isinstance(f2_value, str)
            and 'C2:C25' in formula_normalized
        )
        if has_c_range:
            print(f"PASS: Component 3 — CORREL formula references C2:C25 (units sold column) (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 3 — Expected C2:C25 in formula, got: {repr(f2_value)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
