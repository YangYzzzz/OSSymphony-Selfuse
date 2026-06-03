"""
Reward Script: Calculate total sales for 'Emily Chen' using SUMIF in G2
Task ID: calc_fmb_sumif_single_007
Domain: libreoffice_calc
Scoring:
  - Component 1: G2 contains a SUMIF formula (0.5 pts)
  - Component 2: SUMIF formula uses correct range, criteria, and sum_range (0.5 pts)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_sumif_single_007'


def verify_task(file_path):
    """
    Verify that G2 contains a correct SUMIF formula for Emily Chen's total sales.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Transactions' sheet must exist
    if 'Transactions' not in wb.sheetnames:
        print("FAIL: Sheet 'Transactions' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Transactions']

    # Component 1: G2 contains a SUMIF formula (0.5 points)
    # The task requires placing a SUMIF formula in cell G2.
    # Initial file has G2 = None; golden file has G2 = '=SUMIF(...)'
    try:
        g2_value = ws.cell(row=2, column=7).value
        if g2_value is not None and isinstance(g2_value, str) and 'SUMIF' in g2_value.upper():
            print(f"PASS: Component 1 — G2 contains a SUMIF formula: {repr(g2_value)} (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — Expected SUMIF formula in G2, found: {repr(g2_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not read G2: {e}")

    # Component 2: SUMIF formula has correct arguments (0.5 points)
    # The formula must:
    #   - range: C2:C201 (or equivalent that covers all 200 data rows in column C)
    #   - criteria: "Emily Chen" (case-insensitive match is fine in Excel/Calc, but formula text must include name)
    #   - sum_range: D2:D201 (or equivalent that covers all 200 data rows in column D)
    # Expected: =SUMIF(C2:C201,"Emily Chen",D2:D201)
    try:
        g2_value = ws.cell(row=2, column=7).value
        if g2_value is not None and isinstance(g2_value, str):
            formula_normalized = g2_value.upper().replace(' ', '')
            # Check that the criteria includes "EMILY CHEN"
            has_emily_chen = 'EMILY CHEN' in g2_value.upper()
            # Check range covers C column rows 2-201 (at least C2:C201)
            has_c_range = bool(re.search(r'C2:C20[1-9]', g2_value.upper()) or
                               re.search(r'C2:C2[0-9]{2}', g2_value.upper()) or
                               'C:C' in g2_value.upper())
            # Check sum_range covers D column rows 2-201 (at least D2:D201)
            has_d_range = bool(re.search(r'D2:D20[1-9]', g2_value.upper()) or
                               re.search(r'D2:D2[0-9]{2}', g2_value.upper()) or
                               'D:D' in g2_value.upper())

            if has_emily_chen and has_c_range and has_d_range:
                print(f"PASS: Component 2 — SUMIF formula has correct arguments: "
                      f"criteria='Emily Chen', C-range={has_c_range}, D-range={has_d_range} (0.5 pts)")
                total_score += 0.5
            else:
                print(f"FAIL: Component 2 — SUMIF formula arguments incorrect. "
                      f"has_emily_chen={has_emily_chen}, has_c_range={has_c_range}, has_d_range={has_d_range}. "
                      f"Formula: {repr(g2_value)}")
        else:
            print(f"FAIL: Component 2 — G2 has no formula to check arguments for: {repr(g2_value)}")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not validate SUMIF arguments: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
