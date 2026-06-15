"""
Initial Setup: Sales transactions spreadsheet for SUMIF task
Task ID: calc_fmb_sumif_single_007
Domain: libreoffice_calc
"""

import openpyxl
from datetime import date, timedelta
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_sumif_single_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    random.seed(42)  # reproducible

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Transactions'

    # --- Row 1: Headers ---
    headers = ['Transaction ID', 'Date', 'Salesperson', 'Sale Amount', 'Region', 'Product']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # --- Define salespeople and their transaction counts ---
    # Emily Chen: 47 transactions, total must be 284750
    # James Wilson: 38 transactions
    # Sarah Park: 42 transactions
    # David Lee: 35 transactions
    # Maria Gonzalez: 38 transactions
    # Total: 200 transactions

    salesperson_list = (
        ['Emily Chen'] * 47 +
        ['James Wilson'] * 38 +
        ['Sarah Park'] * 42 +
        ['David Lee'] * 35 +
        ['Maria Gonzalez'] * 38
    )
    random.shuffle(salesperson_list)

    regions = ['North', 'South', 'East', 'West', 'Central']
    products = [
        'Enterprise License', 'Support Package', 'Cloud Subscription',
        'Hardware Bundle', 'Consulting Hours', 'Training Package',
        'Data Analytics Suite', 'Security Module', 'Integration Services',
        'Professional Services'
    ]

    # Generate Emily Chen's sales that sum to exactly 284750
    # 47 transactions, amounts between 500 and 15000
    # We'll generate 46 random amounts and set the 47th to make the sum correct
    emily_amounts = []
    target_sum = 284750
    for i in range(46):
        amt = random.choice([
            1500, 2000, 2500, 3000, 3500, 4000, 4500, 5000,
            5500, 6000, 6500, 7000, 7500, 8000, 9000, 10000,
            11000, 12000
        ])
        emily_amounts.append(amt)
    # Calculate last amount to hit exact target
    last_amt = target_sum - sum(emily_amounts)
    # Ensure last_amt is within valid range [500, 15000]
    if last_amt < 500:
        # Adjust: reduce some amounts
        diff = 500 - last_amt
        for i in range(len(emily_amounts)):
            if emily_amounts[i] - diff >= 500:
                emily_amounts[i] -= diff
                last_amt = 500
                break
    elif last_amt > 15000:
        # Adjust: increase some amounts
        diff = last_amt - 15000
        for i in range(len(emily_amounts)):
            if emily_amounts[i] + diff <= 15000:
                emily_amounts[i] += diff
                last_amt = 15000
                break
    emily_amounts.append(last_amt)

    # Verify Emily's total
    assert sum(emily_amounts) == 284750, f"Emily's sum is {sum(emily_amounts)}, expected 284750"

    # Generate amounts for other salespeople
    other_amounts = {}
    for sp in ['James Wilson', 'Sarah Park', 'David Lee', 'Maria Gonzalez']:
        count = salesperson_list.count(sp)
        amounts = [random.randint(1, 30) * 500 for _ in range(count)]
        # Clamp to [500, 15000]
        amounts = [max(500, min(15000, a)) for a in amounts]
        other_amounts[sp] = amounts

    # Build amount lookup per salesperson (pop from their list)
    emily_idx = 0
    other_idx = {sp: 0 for sp in other_amounts}

    # Generate start date
    base_date = date(2024, 1, 2)

    # Write 200 rows
    for row_num, sp in enumerate(salesperson_list, 2):
        tx_id = f'TXN-{row_num - 1:04d}'
        tx_date = base_date + timedelta(days=(row_num - 2))
        region = random.choice(regions)
        product = random.choice(products)

        if sp == 'Emily Chen':
            amount = emily_amounts[emily_idx]
            emily_idx += 1
        else:
            idx = other_idx[sp]
            amount = other_amounts[sp][idx]
            other_idx[sp] += 1

        ws.cell(row=row_num, column=1, value=tx_id)
        ws.cell(row=row_num, column=2, value=tx_date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=3, value=sp)
        ws.cell(row=row_num, column=4, value=amount)
        ws.cell(row=row_num, column=5, value=region)
        ws.cell(row=row_num, column=6, value=product)

    # --- F2: label for the target ---
    # NOTE: F2 is the label 'Emily Chen Total', G2 is empty (task target)
    ws.cell(row=2, column=6, value='Emily Chen Total')
    # G2 must be empty - do NOT write anything there

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Transactions')
    print(f'  Rows: 201 (1 header + 200 data)')
    print(f'  F2: Emily Chen Total (label)')
    print(f'  G2: empty (target cell)')

create_initial()
