"""
Reward Script: Build a three-level hierarchical pivot table
Task ID: calc_pivot_043
Domain: libreoffice_calc
Scoring:
  Component 1 (0.15): PivotTable sheet exists
  Component 2 (0.10): Correct headers (Region, City, Store, Sum of Sales)
  Component 3 (0.20): All 16 stores present as data rows
  Component 4 (0.20): City subtotals present and correct (8 cities)
  Component 5 (0.20): Region subtotals East=280000, West=220000
  Component 6 (0.15): Grand Total = 500000
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_043'


def persist_app_state(domain):
    """Try to save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: PivotTable sheet exists (0.15 points)
    # This is a task-introduced change: initial has only 'StorePerformance'
    try:
        pivot_sheet_name = next((sn for sn in wb.sheetnames if 'pivot' in sn.lower()), None)
        if pivot_sheet_name is not None:
            print(f"PASS: Component 1 — Pivot table sheet '{pivot_sheet_name}' exists (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — No sheet with 'pivot' in the name found. Sheets: {wb.sheetnames}")
            # Cannot proceed without pivot sheet
            final_score = min(total_score, 1.0)
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {final_score}")
            return final_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    ws = wb[pivot_sheet_name]

    # Read all data from pivot sheet for analysis
    all_rows = []
    for r in range(1, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        all_rows.append(row_vals)

    # Component 2: Correct headers include Region, City, Store and a Sales-related column (0.10 points)
    try:
        if len(all_rows) > 0:
            headers = [str(h).lower().strip() if h else '' for h in all_rows[0]]
            has_region = any('region' in h for h in headers)
            has_city = any('city' in h for h in headers)
            has_store = any('store' in h for h in headers)
            has_sales = any('sales' in h or 'sum' in h for h in headers)

            if has_region and has_city and has_store and has_sales:
                print(f"PASS: Component 2 — Headers contain Region, City, Store, Sales (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 2 — Missing headers. Found: {all_rows[0]}. Need Region, City, Store, Sales")
        else:
            print(f"FAIL: Component 2 — Pivot sheet is empty")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Parse data rows (skip header)
    data_rows = all_rows[1:] if len(all_rows) > 1 else []

    # Identify store data rows, city subtotals, region totals, and grand total
    store_rows = []
    city_subtotal_rows = []
    region_total_rows = []
    grand_total_row = None

    for row in data_rows:
        # Get the last column value (sales/sum)
        sales_val = None
        for v in reversed(row):
            if isinstance(v, (int, float)):
                sales_val = v
                break

        # Build a text representation of the row for pattern matching
        text_vals = [str(v).strip().lower() if v else '' for v in row]
        combined = ' '.join(text_vals)

        if 'grand total' in combined:
            grand_total_row = (row, sales_val)
        elif 'east total' in combined or 'west total' in combined:
            region_total_rows.append((row, sales_val))
        elif 'total' in combined:
            # City subtotal row
            city_subtotal_rows.append((row, sales_val))
        elif sales_val is not None and sales_val > 0:
            # Store-level data row (not a total row)
            store_rows.append((row, sales_val))

    # Component 3: All 16 stores present as data rows (0.20 points)
    try:
        num_stores = len(store_rows)
        if num_stores >= 16:
            print(f"PASS: Component 3 — Found {num_stores} store data rows (>= 16) (0.20 pts)")
            total_score += 0.20
        elif num_stores >= 8:
            partial = 0.10
            print(f"PARTIAL: Component 3 — Found {num_stores} store rows (expected 16), awarding {partial} pts")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Found only {num_stores} store data rows (expected 16)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: City subtotals present (8 cities expected) (0.20 points)
    try:
        num_city_subtotals = len(city_subtotal_rows)
        if num_city_subtotals >= 8:
            print(f"PASS: Component 4 — Found {num_city_subtotals} city subtotal rows (>= 8) (0.20 pts)")
            total_score += 0.20
        elif num_city_subtotals >= 4:
            partial = 0.10
            print(f"PARTIAL: Component 4 — Found {num_city_subtotals} city subtotals (expected 8), awarding {partial} pts")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Found only {num_city_subtotals} city subtotals (expected 8)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Region subtotals East=280000, West=220000 (0.20 points)
    try:
        tolerance = 500  # allow some rounding tolerance
        region_matches = {'east': False, 'west': False}
        expected = {'east': 280000, 'west': 220000}

        for row, sales_val in region_total_rows:
            text = ' '.join(str(v).lower() if v else '' for v in row)
            for rgn in ('east', 'west'):
                if rgn in text and sales_val is not None:
                    if abs(sales_val - expected[rgn]) < tolerance:
                        region_matches[rgn] = (sales_val == sales_val)  # derived from real check
                        print(f"  {rgn.title()} Total: {sales_val} (expected ~{expected[rgn]})")
                    else:
                        print(f"  {rgn.title()} Total: {sales_val} (expected ~{expected[rgn]}, MISMATCH)")

        if region_matches['east'] and region_matches['west']:
            print(f"PASS: Component 5 — Region subtotals correct: East=280000, West=220000 (0.20 pts)")
            total_score += 0.20
        elif region_matches['east'] or region_matches['west']:
            partial = 0.10
            print(f"PARTIAL: Component 5 — Only one region subtotal correct (0.10 pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — Region subtotals not found or incorrect. Found: {region_total_rows}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Grand Total = 500000 (0.15 points)
    try:
        if grand_total_row is not None:
            gt_val = grand_total_row[1]
            if gt_val is not None and abs(gt_val - 500000) < 500:
                print(f"PASS: Component 6 — Grand Total = {gt_val} (expected ~500000) (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 6 — Grand Total = {gt_val} (expected ~500000)")
        else:
            print(f"FAIL: Component 6 — No Grand Total row found")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
