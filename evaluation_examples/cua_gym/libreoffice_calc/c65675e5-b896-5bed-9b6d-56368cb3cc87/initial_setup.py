"""
Initial Setup: Build a three-level hierarchical pivot table
Task ID: calc_pivot_043
Domain: libreoffice_calc

Creates StorePerformance sheet with 500 transaction rows:
  TxnID, Region, City, Store, Date, Sales
  East subtotal=280000, West subtotal=220000, Grand total=500000
"""

import os
import random
import shlex
import subprocess
import time
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_043'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

# ---------- geography hierarchy ----------
HIERARCHY = {
    'East': {
        'New York': ['NYC-Downtown', 'NYC-Midtown'],
        'Boston': ['Boston-Back Bay', 'Boston-Cambridge'],
        'Philadelphia': ['Philly-Center', 'Philly-University'],
        'Miami': ['Miami-Beach', 'Miami-Brickell'],
    },
    'West': {
        'Los Angeles': ['LA-Hollywood', 'LA-Santa Monica'],
        'San Francisco': ['SF-Union Square', 'SF-Marina'],
        'Seattle': ['Seattle-Downtown', 'Seattle-Capitol Hill'],
        'Denver': ['Denver-LoDo', 'Denver-Cherry Creek'],
    },
}


def launch_gui(command: str, delay_sec: float = 1.0):
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def generate_sales_data():
    """Generate 500 rows with controlled totals: East=280000, West=220000."""
    random.seed(42)

    # Build flat store list with region info
    stores = []
    for region, cities in HIERARCHY.items():
        for city, store_list in cities.items():
            for store in store_list:
                stores.append((region, city, store))

    east_stores = [s for s in stores if s[0] == 'East']
    west_stores = [s for s in stores if s[0] == 'West']

    # Generate 250 East rows summing to 280000
    east_rows = _generate_region_rows(east_stores, 250, 280000)
    # Generate 250 West rows summing to 220000
    west_rows = _generate_region_rows(west_stores, 250, 220000)

    all_rows = east_rows + west_rows
    random.shuffle(all_rows)

    # Assign TxnIDs and dates
    base_date = date(2025, 1, 1)
    result = []
    for i, (region, city, store, sales) in enumerate(all_rows, 1):
        txn_date = base_date + timedelta(days=random.randint(0, 364))
        result.append((i, region, city, store, txn_date, sales))
    return result


def _generate_region_rows(store_list, n_rows, target_total):
    """Generate n_rows for given stores summing exactly to target_total."""
    rows = []
    # Random amounts first
    raw = [random.randint(200, 1500) for _ in range(n_rows)]
    raw_sum = sum(raw)
    # Scale to target
    scaled = [round(v * target_total / raw_sum, 2) for v in raw]
    # Fix rounding error on last element
    diff = round(target_total - sum(scaled), 2)
    scaled[-1] = round(scaled[-1] + diff, 2)

    for idx, amt in enumerate(scaled):
        store_info = store_list[idx % len(store_list)]
        rows.append((store_info[0], store_info[1], store_info[2], amt))
    return rows


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'StorePerformance'

    # --- Header row ---
    headers = ['TxnID', 'Region', 'City', 'Store', 'Date', 'Sales']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    thin = Side(style="thin", color="000000")
    header_border = Border(bottom=Side(style="medium", color="000000"),
                           left=thin, right=thin, top=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # --- Data rows ---
    data = generate_sales_data()
    for r, (txn_id, region, city, store, txn_date, sales) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=txn_id)
        ws.cell(row=r, column=2, value=region)
        ws.cell(row=r, column=3, value=city)
        ws.cell(row=r, column=4, value=store)
        ws.cell(row=r, column=5, value=txn_date)
        ws.cell(row=r, column=5).number_format = 'yyyy-mm-dd'
        ws.cell(row=r, column=6, value=sales)
        ws.cell(row=r, column=6).number_format = '#,##0.00'

    # --- Column widths ---
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 24
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14

    # Freeze header
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Verify totals
    east_total = sum(row[5] for row in data if row[1] == 'East')
    west_total = sum(row[5] for row in data if row[1] == 'West')
    print(f'East total: {east_total}, West total: {west_total}, Grand total: {east_total + west_total}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
