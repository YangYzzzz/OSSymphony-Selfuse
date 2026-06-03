"""
Initial Setup: Create a Market Share spreadsheet with region/revenue data
Task ID: calc_gg1_015
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_015'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Market Share sheet ---
    ws = wb.active
    ws.title = 'Market Share'

    # Headers
    ws.cell(row=1, column=1, value='Region')
    ws.cell(row=1, column=2, value='Revenue')

    # Data: 5 regions with realistic revenue values
    data = [
        ['North', 245800],
        ['South', 187500],
        ['East', 312400],
        ['West', 156900],
        ['International', 198300],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 14

    # Format revenue as currency
    for r in range(2, 7):
        ws.cell(row=r, column=2).number_format = '$#,##0'

    # NO charts — the task asks the agent to create the chart
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
