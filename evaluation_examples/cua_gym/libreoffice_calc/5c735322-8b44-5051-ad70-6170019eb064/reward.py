"""
Reward Script: HR Salary Percentile Ranking
Task ID: calc_hr_salary_percentile_037
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Column E has PERCENTRANK formulas in E2:E118      (0.40 pts)
  Component 2: Column E cells E2:E118 formatted as percentage    (0.20 pts)
  Component 3: Conditional formatting applied with yellow fill    (0.30 pts)
  Component 4: CF range covers A2:E118 and condition is E<0.25   (0.10 pts)
  Total: 1.00

Verification strategy:
  - Load golden/initial .xlsx file with openpyxl
  - Check column E formulas match the IFERROR(COUNTIFS/COUNTIF) pattern
  - Check number_format is '0%' on data rows
  - Check conditional_formatting has at least one rule referencing E column < 0.25
    with a yellow (FFFFFF00) fill
  - Column A-D integrity is checked as a precondition gate only
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_salary_percentile_037'

# Expected formula pattern for column E
# =IFERROR(COUNTIFS($C$2:$C$118,C2,$D$2:$D$118,"<"&D2)/COUNTIF($C$2:$C$118,C2),0)
# The row reference in the formula varies per row (C2, D2 → C3, D3, etc.)
FORMULA_PATTERN = re.compile(
    r'=IFERROR\s*\(\s*COUNTIFS\s*\(\s*\$C\$2:\$C\$118\s*,\s*C\d+\s*,\s*\$D\$2:\$D\$118\s*,\s*"<"\s*&\s*D\d+\s*\)\s*/\s*COUNTIF\s*\(\s*\$C\$2:\$C\$118\s*,\s*C\d+\s*\)\s*,\s*0\s*\)',
    re.IGNORECASE
)

YELLOW_COLOR = 'FFFFFF00'
YELLOW_ALT = 'FFFF00'  # some tools may store 6-char


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # --- Precondition gate: load workbook ---
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # --- Precondition gate: correct sheet exists ---
    if 'Compensation' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Compensation' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Compensation']

    # --- Precondition gate: columns A-D not modified (spot-check header row) ---
    expected_headers = {1: 'Emp ID', 2: 'Name', 3: 'Department', 4: 'Salary', 5: 'Dept Percentile'}
    headers_ok = True
    for col, name in expected_headers.items():
        actual = ws.cell(1, col).value
        if actual != name:
            print(f"CRITICAL: Header mismatch at col {col}: expected '{name}', found '{actual}'")
            headers_ok = False
    if not headers_ok:
        print("REWARD: 0.0")
        return 0.0

    # -------------------------------------------------------------------------
    # Component 1: Column E has PERCENTRANK formulas in E2:E118 (0.40 points)
    # The task requires an IFERROR(COUNTIFS/COUNTIF) formula in every data row.
    # In the initial file, E2:E118 are all None.
    # -------------------------------------------------------------------------
    try:
        formula_count = 0
        formula_correct = 0
        total_rows = 117  # rows 2-118

        for row in range(2, 119):
            val = ws.cell(row, 5).value
            if val is not None:
                formula_count += 1
                if isinstance(val, str) and FORMULA_PATTERN.match(val.strip()):
                    formula_correct += 1

        if formula_count == 0:
            print(f"FAIL: Component 1 — Column E is empty (no formulas found)")
        elif formula_correct == total_rows:
            print(f"PASS: Component 1 — All {total_rows} rows have correct PERCENTRANK formula (0.40 pts)")
            total_score += 0.40
        elif formula_correct >= total_rows * 0.9:
            # Partial credit: >90% correct formulas
            partial = round(0.40 * formula_correct / total_rows, 2)
            print(f"PARTIAL: Component 1 — {formula_correct}/{total_rows} rows have correct formula ({partial} pts)")
            total_score += partial
        else:
            # Has formulas but they don't match expected pattern
            sample_val = ws.cell(2, 5).value
            print(f"FAIL: Component 1 — Only {formula_correct}/{total_rows} formulas match pattern; sample E2={repr(sample_val)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Column E cells E2:E118 formatted as percentage '0%' (0.20 pts)
    # In the initial file, number_format for these cells is 'General'.
    # -------------------------------------------------------------------------
    try:
        pct_count = 0
        total_rows = 117  # rows 2-118

        for row in range(2, 119):
            cell = ws.cell(row, 5)
            nf = cell.number_format
            if nf in ('0%', '0.0%', '0.00%', '#,##0%'):
                pct_count += 1

        if pct_count == total_rows:
            print(f"PASS: Component 2 — All {total_rows} Column E rows have percentage number format (0.20 pts)")
            total_score += 0.20
        elif pct_count >= total_rows * 0.9:
            partial = round(0.20 * pct_count / total_rows, 2)
            print(f"PARTIAL: Component 2 — {pct_count}/{total_rows} rows have percentage format ({partial} pts)")
            total_score += partial
        else:
            sample_nf = ws.cell(2, 5).number_format
            print(f"FAIL: Component 2 — Only {pct_count}/{total_rows} cells have percentage format; E2 format={repr(sample_nf)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Conditional formatting with yellow (#FFFF00 / FFFFFF00) fill
    #              applied to the data area (0.30 points)
    # In the initial file, there is NO conditional formatting.
    # -------------------------------------------------------------------------
    try:
        cf_rules = list(ws.conditional_formatting)
        found_yellow_cf = False
        cf_count = len(cf_rules)

        for cf_range in cf_rules:
            for rule in cf_range.rules:
                # Check for yellow fill in the rule's differential style
                if rule.dxf and rule.dxf.fill:
                    try:
                        fg = rule.dxf.fill.fgColor.rgb
                        # Accept FFFFFF00 (8-char ARGB yellow) or FFFF00 (6-char)
                        if fg in (YELLOW_COLOR, YELLOW_ALT, '00FFFF00'):
                            found_yellow_cf = True
                            print(f"  CF rule found: range={cf_range}, fgColor={fg}, formula={rule.formula}")
                            break
                    except Exception:
                        pass
            if found_yellow_cf:
                break

        if found_yellow_cf:
            print(f"PASS: Component 3 — Conditional formatting with yellow fill found (0.30 pts)")
            total_score += 0.30
        elif cf_count > 0:
            print(f"FAIL: Component 3 — {cf_count} CF rule(s) found but none with yellow fill")
        else:
            print(f"FAIL: Component 3 — No conditional formatting rules found on sheet")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: CF range covers A2:E118 AND condition references E < 0.25
    #              (0.10 points — refines Component 3)
    # -------------------------------------------------------------------------
    try:
        cf_rules = list(ws.conditional_formatting)
        correct_range = False
        correct_condition = False

        for cf_range in cf_rules:
            # Check range covers at least A2:E118
            range_str = str(cf_range).strip()
            # Accept A2:E118 or larger ranges
            if 'A2:E118' in range_str or 'A2:E' in range_str:
                correct_range = True

            for rule in cf_range.rules:
                # Check condition formula mentions E column < 0.25
                if rule.formula:
                    for f in rule.formula:
                        f_norm = str(f).replace(' ', '').upper()
                        # Looking for $E2<0.25 or E2<0.25 or E<0.25
                        if ('<0.25' in f_norm or '<0.25' in f_norm) and 'E' in f_norm:
                            correct_condition = True

        if correct_range and correct_condition:
            print(f"PASS: Component 4 — CF range A2:E118 with condition E<0.25 verified (0.10 pts)")
            total_score += 0.10
        elif correct_range:
            print(f"FAIL: Component 4 — CF range correct but condition does not reference E<0.25")
        elif correct_condition:
            print(f"FAIL: Component 4 — CF condition correct but range does not cover A2:E118 (found: {range_str})")
        else:
            print(f"FAIL: Component 4 — CF range and condition not matching expected (A2:E118, $E<0.25)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -------------------------------------------------------------------------
    # Final score
    # -------------------------------------------------------------------------
    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
