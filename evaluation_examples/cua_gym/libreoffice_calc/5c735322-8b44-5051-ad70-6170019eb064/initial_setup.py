"""
Initial Setup: HR Salary Compensation Review Spreadsheet
Task ID: calc_hr_salary_percentile_037
Domain: libreoffice_calc

Creates a Compensation sheet with 117 employee records (rows 2-118),
columns: Emp ID, Name, Department, Salary, Dept Percentile (empty).
No formulas, no conditional formatting in column E — that is the task.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_salary_percentile_037'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Compensation'

    # ---- Header Row ----
    headers = ['Emp ID', 'Name', 'Department', 'Salary', 'Dept Percentile']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # ---- Employee Data ----
    # Departments: Engineering, Marketing, Finance, HR, Operations, Sales, Legal
    departments = [
        'Engineering', 'Engineering', 'Engineering', 'Engineering', 'Engineering',
        'Engineering', 'Engineering', 'Engineering', 'Engineering', 'Engineering',
        'Engineering', 'Engineering', 'Engineering', 'Engineering', 'Engineering',
        'Engineering', 'Engineering', 'Engineering', 'Engineering', 'Engineering',
        'Marketing', 'Marketing', 'Marketing', 'Marketing', 'Marketing',
        'Marketing', 'Marketing', 'Marketing', 'Marketing', 'Marketing',
        'Marketing', 'Marketing', 'Marketing', 'Marketing', 'Marketing',
        'Marketing', 'Marketing',
        'Finance', 'Finance', 'Finance', 'Finance', 'Finance',
        'Finance', 'Finance', 'Finance', 'Finance', 'Finance',
        'Finance', 'Finance', 'Finance', 'Finance', 'Finance',
        'Finance', 'Finance',
        'HR', 'HR', 'HR', 'HR', 'HR',
        'HR', 'HR', 'HR', 'HR', 'HR',
        'HR', 'HR', 'HR', 'HR', 'HR',
        'HR', 'HR',
        'Operations', 'Operations', 'Operations', 'Operations', 'Operations',
        'Operations', 'Operations', 'Operations', 'Operations', 'Operations',
        'Operations', 'Operations', 'Operations', 'Operations', 'Operations',
        'Operations', 'Operations',
        'Sales', 'Sales', 'Sales', 'Sales', 'Sales',
        'Sales', 'Sales', 'Sales', 'Sales', 'Sales',
        'Sales', 'Sales', 'Sales', 'Sales', 'Sales',
        'Sales', 'Sales',
        'Legal', 'Legal', 'Legal', 'Legal', 'Legal',
        'Legal', 'Legal', 'Legal', 'Legal', 'Legal',
        'Legal', 'Legal',
    ]

    names = [
        # Engineering (20)
        'Sarah Chen', 'Marcus Johnson', 'Priya Patel', 'Liam O\'Brien', 'Yuki Tanaka',
        'Carlos Mendez', 'Aisha Williams', 'Noah Fischer', 'Mei Lin', 'James Okafor',
        'Elena Vasquez', 'Raj Sharma', 'Olivia Bennett', 'Dmitri Volkov', 'Fatima Al-Hassan',
        'Lucas Dubois', 'Hannah Schmidt', 'Kwame Asante', 'Sophie Larsen', 'Ivan Petrov',
        # Marketing (17)
        'Diana Torres', 'Samuel Park', 'Chloe Martin', 'Antoine Rousseau', 'Layla Hassan',
        'Ben Nakamura', 'Ingrid Johansson', 'Victor Osei', 'Tara Murphy', 'Felix Wagner',
        'Nadia Kowalski', 'Oscar Reyes', 'Amara Diallo', 'Patrick O\'Sullivan', 'Zoe Thompson',
        'Rashid Abdullah', 'Sienna Clarke',
        # Finance (17)
        'Michael Chang', 'Jennifer Liu', 'Robert Kim', 'Amanda Foster', 'David Nguyen',
        'Laura Hernandez', 'Brian Wilson', 'Christina Lee', 'Thomas Brown', 'Michelle Davis',
        'Andrew Miller', 'Stephanie Moore', 'Jonathan Taylor', 'Rebecca Anderson', 'Kevin White',
        'Ashley Jackson', 'Ryan Harris',
        # HR (17)
        'Emma Rodriguez', 'Tyler Martinez', 'Grace Thompson', 'Brandon Hall', 'Madison Lewis',
        'Jordan Walker', 'Taylor Robinson', 'Morgan Young', 'Cameron King', 'Skyler Wright',
        'Riley Scott', 'Avery Green', 'Blake Nelson', 'Quinn Baker', 'Parker Adams',
        'Kendall Carter', 'Casey Mitchell',
        # Operations (17)
        'George Perez', 'Linda Roberts', 'Frank Turner', 'Barbara Phillips', 'Gary Campbell',
        'Sandra Parker', 'Raymond Evans', 'Deborah Edwards', 'Jerry Collins', 'Sharon Stewart',
        'Walter Morris', 'Donna Rogers', 'Harold Reed', 'Ruth Cook', 'Harry Morgan',
        'Helen Bell', 'Clarence Murphy',
        # Sales (17)
        'Chris Rivera', 'Patricia Cooper', 'Steven Richardson', 'Dorothy Cox', 'Mark Howard',
        'Karen Ward', 'Paul Torres', 'Betty Peterson', 'Donald Gray', 'Kimberly Ramirez',
        'Charles James', 'Sandra Watson', 'Jose Brooks', 'Margaret Kelly', 'Daniel Sanders',
        'Angela Price', 'Matthew Bennett',
        # Legal (12)
        'Nicholas Coleman', 'Victoria Jenkins', 'Alexander Powell', 'Natalie Long',
        'Benjamin Patterson', 'Isabella Russell', 'William Simmons', 'Charlotte Foster',
        'Daniel Griffin', 'Abigail Hayes', 'Ethan Myers', 'Sophia Richardson',
    ]

    salaries = [
        # Engineering (20) — range $75k–$145k
        95000, 115000, 82000, 130000, 75000,
        142000, 88000, 105000, 78000, 125000,
        98000, 135000, 70000, 148000, 85000,
        110000, 92000, 145000, 67000, 120000,
        # Marketing (17) — range $58k–$105k
        72000, 89000, 65000, 95000, 58000,
        105000, 78000, 62000, 92000, 68000,
        84000, 55000, 99000, 74000, 81000,
        60000, 96000,
        # Finance (17) — range $72k–$130k
        88000, 115000, 72000, 105000, 130000,
        79000, 96000, 122000, 68000, 108000,
        84000, 92000, 75000, 118000, 85000,
        98000, 128000,
        # HR (17) — range $52k–$92k
        68000, 82000, 57000, 75000, 90000,
        52000, 88000, 63000, 78000, 55000,
        86000, 72000, 59000, 92000, 65000,
        80000, 48000,
        # Operations (17) — range $48k–$88k
        62000, 78000, 55000, 85000, 48000,
        72000, 68000, 58000, 82000, 51000,
        75000, 65000, 88000, 53000, 70000,
        60000, 80000,
        # Sales (17) — range $55k–$115k
        75000, 92000, 62000, 108000, 55000,
        115000, 82000, 68000, 98000, 59000,
        88000, 72000, 105000, 65000, 95000,
        78000, 112000,
        # Legal (12) — range $95k–$165k
        125000, 145000, 98000, 162000,
        110000, 138000, 95000, 155000,
        120000, 148000, 105000, 165000,
    ]

    assert len(departments) == 117, f'Expected 117 rows, got {len(departments)}'
    assert len(names) == 117, f'Expected 117 names, got {len(names)}'
    assert len(salaries) == 117, f'Expected 117 salaries, got {len(salaries)}'

    for i in range(117):
        row = i + 2
        emp_id = f'EMP{1001 + i:04d}'
        ws.cell(row=row, column=1, value=emp_id)
        ws.cell(row=row, column=2, value=names[i])
        ws.cell(row=row, column=3, value=departments[i])
        ws.cell(row=row, column=4, value=salaries[i])
        # Column E (Dept Percentile) intentionally left EMPTY
        # No formula, no value — this is what the agent must add

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 18

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheets: {wb.sheetnames}')
    print(f'  Rows: {ws.max_row} (1 header + 117 data rows)')
    print(f'  Col E (Dept Percentile): EMPTY (no formulas, no values)')

create_initial()
