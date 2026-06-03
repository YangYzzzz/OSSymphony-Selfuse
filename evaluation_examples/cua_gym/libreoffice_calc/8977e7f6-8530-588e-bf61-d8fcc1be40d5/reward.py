"""
Reward Script: Calculate need-based financial aid packages
Task ID: calc_edu_financial_aid_need_051
Domain: libreoffice_calc

Scoring:
  Component 1: EFC formula in column C (IFS income bracket formula) — 0.40 pts
  Component 2: Financial Need formula in column E (=MAX(0,D-C)) — 0.20 pts
  Component 3a: Grant formula in column F (=MIN(E,6000)) — 0.15 pts
  Component 3b: Sub Loan formula in column G (=MIN(MAX(0,E-F),5500)) — 0.10 pts
  Component 3c: Work Study formula in column H (=MIN(MAX(0,E-F-G),3000)) — 0.10 pts
  Component 4: Unmet Need formula in column I (=MAX(0,E-F-G-H)) — 0.05 pts
  Total: 1.00

NOTE: Currency formatting is a precondition (pre-existing in initial file)
and is NOT scored here.
"""

import os
import re
import openpyxl
from openpyxl.utils import get_column_letter

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_edu_financial_aid_need_051'
SHEET_NAME = 'AidPackaging'
DATA_ROWS = range(2, 62)  # rows 2-61 (60 students)


def normalize_formula(formula):
    """Normalize a formula for comparison: uppercase, remove spaces."""
    if not isinstance(formula, str):
        return ''
    return formula.upper().replace(' ', '')


def check_ifs_efc_formula(formula_str):
    """
    Check if formula is a valid IFS formula for EFC calculation.
    Expected pattern (any row): IFS with 4 conditions matching:
      income < 30000 -> 0
      income <= 60000 -> income * 0.10
      income <= 100000 -> income * 0.15
      income > 100000 -> income * 0.22
    """
    if not isinstance(formula_str, str):
        return False
    f = normalize_formula(formula_str)
    if not f.startswith('=IFS('):
        return False
    # Check for the key bracket thresholds and multipliers
    checks = [
        r'30000',
        r'60000',
        r'0\.10',
        r'100000',
        r'0\.15',
        r'0\.22',
    ]
    for pattern in checks:
        if not re.search(pattern, f):
            return False
    return True


def check_max_need_formula(formula_str, row):
    """
    Check if formula is =MAX(0,D{row}-C{row}) or equivalent.
    """
    if not isinstance(formula_str, str):
        return False
    f = normalize_formula(formula_str)
    expected = f'=MAX(0,D{row}-C{row})'
    return f == expected


def check_grant_formula(formula_str, row):
    """Check if formula is =MIN(E{row},6000)."""
    if not isinstance(formula_str, str):
        return False
    f = normalize_formula(formula_str)
    expected = f'=MIN(E{row},6000)'
    return f == expected


def check_subloan_formula(formula_str, row):
    """Check if formula is =MIN(MAX(0,E{row}-F{row}),5500)."""
    if not isinstance(formula_str, str):
        return False
    f = normalize_formula(formula_str)
    expected = f'=MIN(MAX(0,E{row}-F{row}),5500)'
    return f == expected


def check_workstudy_formula(formula_str, row):
    """Check if formula is =MIN(MAX(0,E{row}-F{row}-G{row}),3000)."""
    if not isinstance(formula_str, str):
        return False
    f = normalize_formula(formula_str)
    expected = f'=MIN(MAX(0,E{row}-F{row}-G{row}),3000)'
    return f == expected


def check_unmet_formula(formula_str, row):
    """Check if formula is =MAX(0,E{row}-F{row}-G{row}-H{row})."""
    if not isinstance(formula_str, str):
        return False
    f = normalize_formula(formula_str)
    expected = f'=MAX(0,E{row}-F{row}-G{row}-H{row})'
    return f == expected


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet exists (precondition gate)
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # -----------------------------------------------------------------------
    # Component 1: EFC formula in column C — IFS income bracket formula (0.40 pts)
    # All 60 rows (2-61) should have the IFS EFC formula with correct brackets.
    # This FAILS on initial (all None) and PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        efc_rows_ok = 0
        efc_rows_missing = []
        for row in DATA_ROWS:
            val = ws.cell(row=row, column=3).value  # column C
            if check_ifs_efc_formula(val):
                efc_rows_ok += 1
            else:
                efc_rows_missing.append((row, val))

        partial_1 = 0.40 if efc_rows_ok == 60 else (round(0.40 * efc_rows_ok / 60, 4) if efc_rows_ok >= 30 else 0.0)
        if partial_1 > 0.0:
            total_score += partial_1
            if efc_rows_ok == 60:
                print(f"PASS: Component 1 — EFC IFS formula present in all 60 rows of column C (0.40 pts)")
            else:
                print(f"PARTIAL: Component 1 — EFC IFS formula in {efc_rows_ok}/60 rows of column C ({partial_1} pts)")
                print(f"  Missing examples: {efc_rows_missing[:3]}")
        else:
            print(f"FAIL: Component 1 — EFC IFS formula only in {efc_rows_ok}/60 rows of column C (0.0 pts)")
            if efc_rows_missing:
                print(f"  First missing: row {efc_rows_missing[0][0]}, value={repr(efc_rows_missing[0][1])}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Financial Need formula in column E — =MAX(0,D{row}-C{row}) (0.20 pts)
    # This FAILS on initial (all None) and PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        need_rows_ok = 0
        need_rows_missing = []
        for row in DATA_ROWS:
            val = ws.cell(row=row, column=5).value  # column E
            if check_max_need_formula(val, row):
                need_rows_ok += 1
            else:
                need_rows_missing.append((row, val))

        partial_2 = 0.20 if need_rows_ok == 60 else (round(0.20 * need_rows_ok / 60, 4) if need_rows_ok >= 30 else 0.0)
        if partial_2 > 0.0:
            total_score += partial_2
            if need_rows_ok == 60:
                print(f"PASS: Component 2 — Financial Need formula present in all 60 rows of column E (0.20 pts)")
            else:
                print(f"PARTIAL: Component 2 — Financial Need formula in {need_rows_ok}/60 rows of column E ({partial_2} pts)")
                print(f"  Missing examples: {need_rows_missing[:3]}")
        else:
            print(f"FAIL: Component 2 — Financial Need formula only in {need_rows_ok}/60 rows of column E (0.0 pts)")
            if need_rows_missing:
                print(f"  First missing: row {need_rows_missing[0][0]}, value={repr(need_rows_missing[0][1])}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3a: Grant formula in column F — =MIN(E{row},6000) (0.15 pts)
    # This FAILS on initial (all None) and PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        grant_ok = 0
        grant_missing = []
        for row in DATA_ROWS:
            val = ws.cell(row=row, column=6).value  # column F
            if check_grant_formula(val, row):
                grant_ok += 1
            else:
                grant_missing.append((row, val))

        partial_3a = 0.15 if grant_ok == 60 else (round(0.15 * grant_ok / 60, 4) if grant_ok >= 30 else 0.0)
        if partial_3a > 0.0:
            total_score += partial_3a
            if grant_ok == 60:
                print(f"PASS: Component 3a — Grant formula present in all 60 rows of column F (0.15 pts)")
            else:
                print(f"PARTIAL: Component 3a — Grant formula in {grant_ok}/60 rows ({partial_3a} pts)")
        else:
            print(f"FAIL: Component 3a — Grant formula only in {grant_ok}/60 rows (0.0 pts)")
            if grant_missing:
                print(f"  First missing: row {grant_missing[0][0]}, value={repr(grant_missing[0][1])}")
    except Exception as e:
        print(f"ERROR: Component 3a — {e}")

    # -----------------------------------------------------------------------
    # Component 3b: Sub Loan formula in column G — =MIN(MAX(0,E{row}-F{row}),5500) (0.10 pts)
    # This FAILS on initial (all None) and PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        subloan_ok = 0
        subloan_missing = []
        for row in DATA_ROWS:
            val = ws.cell(row=row, column=7).value  # column G
            if check_subloan_formula(val, row):
                subloan_ok += 1
            else:
                subloan_missing.append((row, val))

        partial_3b = 0.10 if subloan_ok == 60 else (round(0.10 * subloan_ok / 60, 4) if subloan_ok >= 30 else 0.0)
        if partial_3b > 0.0:
            total_score += partial_3b
            if subloan_ok == 60:
                print(f"PASS: Component 3b — Sub Loan formula present in all 60 rows of column G (0.10 pts)")
            else:
                print(f"PARTIAL: Component 3b — Sub Loan formula in {subloan_ok}/60 rows ({partial_3b} pts)")
        else:
            print(f"FAIL: Component 3b — Sub Loan formula only in {subloan_ok}/60 rows (0.0 pts)")
            if subloan_missing:
                print(f"  First missing: row {subloan_missing[0][0]}, value={repr(subloan_missing[0][1])}")
    except Exception as e:
        print(f"ERROR: Component 3b — {e}")

    # -----------------------------------------------------------------------
    # Component 3c: Work Study formula in column H — =MIN(MAX(0,E{row}-F{row}-G{row}),3000) (0.10 pts)
    # This FAILS on initial (all None) and PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        ws_ok = 0
        ws_missing = []
        for row in DATA_ROWS:
            val = ws.cell(row=row, column=8).value  # column H
            if check_workstudy_formula(val, row):
                ws_ok += 1
            else:
                ws_missing.append((row, val))

        partial_3c = 0.10 if ws_ok == 60 else (round(0.10 * ws_ok / 60, 4) if ws_ok >= 30 else 0.0)
        if partial_3c > 0.0:
            total_score += partial_3c
            if ws_ok == 60:
                print(f"PASS: Component 3c — Work Study formula present in all 60 rows of column H (0.10 pts)")
            else:
                print(f"PARTIAL: Component 3c — Work Study formula in {ws_ok}/60 rows ({partial_3c} pts)")
        else:
            print(f"FAIL: Component 3c — Work Study formula only in {ws_ok}/60 rows (0.0 pts)")
            if ws_missing:
                print(f"  First missing: row {ws_missing[0][0]}, value={repr(ws_missing[0][1])}")
    except Exception as e:
        print(f"ERROR: Component 3c — {e}")

    # -----------------------------------------------------------------------
    # Component 4: Unmet Need formula in column I — =MAX(0,E{row}-F{row}-G{row}-H{row}) (0.05 pts)
    # This FAILS on initial (all None) and PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        unmet_rows_ok = 0
        unmet_rows_missing = []
        for row in DATA_ROWS:
            val = ws.cell(row=row, column=9).value  # column I
            if check_unmet_formula(val, row):
                unmet_rows_ok += 1
            else:
                unmet_rows_missing.append((row, val))

        partial_4 = 0.05 if unmet_rows_ok == 60 else (round(0.05 * unmet_rows_ok / 60, 4) if unmet_rows_ok >= 30 else 0.0)
        if partial_4 > 0.0:
            total_score += partial_4
            if unmet_rows_ok == 60:
                print(f"PASS: Component 4 — Unmet Need formula present in all 60 rows of column I (0.05 pts)")
            else:
                print(f"PARTIAL: Component 4 — Unmet Need formula in {unmet_rows_ok}/60 rows ({partial_4} pts)")
        else:
            print(f"FAIL: Component 4 — Unmet Need formula only in {unmet_rows_ok}/60 rows (0.0 pts)")
            if unmet_rows_missing:
                print(f"  First missing: row {unmet_rows_missing[0][0]}, value={repr(unmet_rows_missing[0][1])}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
