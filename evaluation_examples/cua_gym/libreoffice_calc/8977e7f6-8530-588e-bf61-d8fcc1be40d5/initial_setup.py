"""
Initial Setup: Financial Aid Packaging Spreadsheet
Task ID: calc_edu_financial_aid_need_051
Domain: libreoffice_calc

Creates a spreadsheet with 60 students' Student ID, Family Income, and COA filled.
Columns C (EFC), E (Financial Need), F (Grant), G (Sub Loan), H (Work Study), I (Unmet Need) are empty.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_financial_aid_need_051'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'AidPackaging'

    # --- Headers ---
    headers = [
        'Student ID', 'Family Income', 'EFC', 'COA',
        'Financial Need', 'Grant', 'Sub Loan', 'Work Study', 'Unmet Need'
    ]
    header_font = Font(name='Calibri', bold=True, size=11)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    ws.row_dimensions[1].height = 20

    # --- Student data: 60 students ---
    # Income distribution covering all 4 EFC brackets
    # <$30k: very low income (EFC=0)
    # $30k-$60k: moderate income (EFC=income*0.10)
    # $60k-$100k: middle income (EFC=income*0.15)
    # >$100k: higher income (EFC=income*0.22)
    # COA ranges from $18,000 to $35,000

    student_data = [
        # Student ID, Family Income, COA
        ('STU-2025-001', 18500, 22000),
        ('STU-2025-002', 24300, 19500),
        ('STU-2025-003', 27800, 28000),
        ('STU-2025-004', 12400, 21000),
        ('STU-2025-005', 9800,  18500),
        ('STU-2025-006', 22100, 25000),
        ('STU-2025-007', 15600, 30000),
        ('STU-2025-008', 28900, 23500),
        ('STU-2025-009', 6500,  20000),
        ('STU-2025-010', 20000, 27000),
        ('STU-2025-011', 32500, 19000),
        ('STU-2025-012', 41200, 22500),
        ('STU-2025-013', 38700, 31000),
        ('STU-2025-014', 55000, 26000),
        ('STU-2025-015', 48300, 20500),
        ('STU-2025-016', 35600, 34000),
        ('STU-2025-017', 43900, 18000),
        ('STU-2025-018', 59200, 29500),
        ('STU-2025-019', 31800, 24000),
        ('STU-2025-020', 46500, 35000),
        ('STU-2025-021', 63000, 21500),
        ('STU-2025-022', 78500, 28500),
        ('STU-2025-023', 71200, 32000),
        ('STU-2025-024', 95600, 19000),
        ('STU-2025-025', 67800, 23000),
        ('STU-2025-026', 82400, 30500),
        ('STU-2025-027', 74100, 27000),
        ('STU-2025-028', 89300, 33500),
        ('STU-2025-029', 61500, 20000),
        ('STU-2025-030', 76800, 25500),
        ('STU-2025-031', 105000, 22000),
        ('STU-2025-032', 128000, 31000),
        ('STU-2025-033', 115600, 18500),
        ('STU-2025-034', 142000, 29000),
        ('STU-2025-035', 108900, 35000),
        ('STU-2025-036', 175000, 24500),
        ('STU-2025-037', 133500, 21000),
        ('STU-2025-038', 162000, 28000),
        ('STU-2025-039', 118400, 32500),
        ('STU-2025-040', 195000, 19500),
        ('STU-2025-041', 14200, 26000),
        ('STU-2025-042', 29500, 20000),
        ('STU-2025-043', 7300,  30000),
        ('STU-2025-044', 23600, 18000),
        ('STU-2025-045', 11800, 34000),
        ('STU-2025-046', 37400, 22000),
        ('STU-2025-047', 52800, 27500),
        ('STU-2025-048', 44100, 31500),
        ('STU-2025-049', 58600, 19000),
        ('STU-2025-050', 33200, 23500),
        ('STU-2025-051', 69400, 29000),
        ('STU-2025-052', 87200, 21500),
        ('STU-2025-053', 72900, 33000),
        ('STU-2025-054', 91500, 25000),
        ('STU-2025-055', 65300, 18000),
        ('STU-2025-056', 121000, 26500),
        ('STU-2025-057', 148000, 35000),
        ('STU-2025-058', 103500, 22500),
        ('STU-2025-059', 167000, 30000),
        ('STU-2025-060', 112800, 28500),
    ]

    # Format for currency columns (B and D in initial file)
    currency_fmt = '$#,##0.00'
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, (stu_id, income, coa) in enumerate(student_data, 2):
        # Col A: Student ID
        cell_a = ws.cell(row=row_idx, column=1, value=stu_id)
        cell_a.alignment = Alignment(horizontal='left')
        cell_a.border = data_border

        # Col B: Family Income (filled)
        cell_b = ws.cell(row=row_idx, column=2, value=income)
        cell_b.number_format = currency_fmt
        cell_b.border = data_border

        # Col C: EFC (intentionally empty — to be calculated by agent)
        cell_c = ws.cell(row=row_idx, column=3, value=None)
        cell_c.number_format = currency_fmt
        cell_c.border = data_border

        # Col D: COA (filled)
        cell_d = ws.cell(row=row_idx, column=4, value=coa)
        cell_d.number_format = currency_fmt
        cell_d.border = data_border

        # Cols E-I: empty (to be calculated by agent)
        for col in range(5, 10):
            cell = ws.cell(row=row_idx, column=col, value=None)
            cell.number_format = currency_fmt
            cell.border = data_border

    # Column widths
    col_widths = {
        'A': 16,  # Student ID
        'B': 16,  # Family Income
        'C': 14,  # EFC
        'D': 14,  # COA
        'E': 16,  # Financial Need
        'F': 10,  # Grant
        'G': 14,  # Sub Loan
        'H': 14,  # Work Study
        'I': 14,  # Unmet Need
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze the header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheets: {wb.sheetnames}')
    print(f'Rows: {ws.max_row}, Columns: {ws.max_column}')


create_initial()
