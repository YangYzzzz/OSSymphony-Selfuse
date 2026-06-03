"""
Initial Setup: Create spreadsheet with two half-year data sheets for pivot table consolidation
Task ID: calc_gcp_088
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_088'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

PRODUCTS = ['Laptop Pro', 'Wireless Mouse', 'USB-C Hub', 'Mechanical Keyboard', 'Monitor 27inch']
REGIONS = ['North America', 'Europe', 'Asia Pacific', 'Latin America']
Q1Q2_MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun']
Q3Q4_MONTHS = ['Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def generate_data(months, seed_offset=0):
    """Generate 200 rows of product/region/revenue data for given months."""
    rng = random.Random(42 + seed_offset)
    data = []
    # Base revenue ranges per product
    base_revenue = {
        'Laptop Pro': (3500, 12000),
        'Wireless Mouse': (80, 220),
        'USB-C Hub': (120, 360),
        'Mechanical Keyboard': (300, 850),
        'Monitor 27inch': (1200, 3200),
    }
    # Region multipliers
    region_mult = {
        'North America': 1.2,
        'Europe': 1.0,
        'Asia Pacific': 0.9,
        'Latin America': 0.8,
    }

    for i in range(200):
        product = PRODUCTS[i % len(PRODUCTS)]
        region = REGIONS[i % len(REGIONS)]
        month = months[i % len(months)]
        low, high = base_revenue[product]
        mult = region_mult[region]
        revenue = round(rng.uniform(low, high) * mult, 2)
        data.append([product, region, month, revenue])

    return data

def create_initial():
    wb = openpyxl.Workbook()

    # Header style
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    white_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # --- Sheet 1: H1Data (Q1-Q2) ---
    ws1 = wb.active
    ws1.title = 'H1Data'

    headers = ['Product', 'Region', 'Month', 'Revenue']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    h1_data = generate_data(Q1Q2_MONTHS, seed_offset=0)
    for r, row_data in enumerate(h1_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            if c == 4:  # Revenue column
                cell.number_format = '$#,##0.00'
            cell.border = thin_border

    # Column widths
    ws1.column_dimensions['A'].width = 22
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 14
    ws1.freeze_panes = 'A2'

    # --- Sheet 2: H2Data (Q3-Q4) ---
    ws2 = wb.create_sheet('H2Data')

    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    h2_data = generate_data(Q3Q4_MONTHS, seed_offset=100)
    for r, row_data in enumerate(h2_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if c == 4:  # Revenue column
                cell.number_format = '$#,##0.00'
            cell.border = thin_border

    # Column widths
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 10
    ws2.column_dimensions['D'].width = 14
    ws2.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Compute and print totals for verification
    h1_total = sum(row[3] for row in h1_data)
    h2_total = sum(row[3] for row in h2_data)
    print(f'H1Data total revenue: ${h1_total:,.2f}')
    print(f'H2Data total revenue: ${h2_total:,.2f}')
    print(f'Combined total revenue: ${h1_total + h2_total:,.2f}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
