"""
Reward Script: Create pivot table consolidating data from H1Data and H2Data sheets
Task ID: calc_gcp_088
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): A new sheet exists beyond the original two (H1Data, H2Data)
  Component 2 (0.25): Pivot structure — Products as rows, Regions as columns
  Component 3 (0.25): Revenue data values are correct (match source sums within tolerance)
  Component 4 (0.25): Grand total is approximately $844,316.62
"""

import os
import collections
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_088'

# Expected products and regions from the source data
EXPECTED_PRODUCTS = {'Laptop Pro', 'Wireless Mouse', 'USB-C Hub', 'Mechanical Keyboard', 'Monitor 27inch'}
EXPECTED_REGIONS = {'North America', 'Europe', 'Asia Pacific', 'Latin America'}
EXPECTED_GRAND_TOTAL = 844316.62


def compute_source_totals(wb):
    """Compute product-by-region revenue sums from H1Data and H2Data source sheets."""
    totals = collections.defaultdict(lambda: collections.defaultdict(float))
    for sn in ['H1Data', 'H2Data']:
        if sn not in wb.sheetnames:
            continue
        src = wb[sn]
        for row in src.iter_rows(min_row=2, max_row=src.max_row, values_only=True):
            if len(row) < 4:
                continue
            product, region, month, revenue = row[0], row[1], row[2], row[3]
            if product and region and revenue is not None:
                try:
                    totals[str(product).strip()][str(region).strip()] += float(revenue)
                except (ValueError, TypeError):
                    pass
    return totals


def find_pivot_sheet(wb):
    """Find the new sheet that is not H1Data or H2Data."""
    for sn in wb.sheetnames:
        if sn not in ('H1Data', 'H2Data'):
            return sn
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: A new sheet exists beyond the original two (0.25 points)
    try:
        pivot_sheet_name = find_pivot_sheet(wb)
        if pivot_sheet_name is not None and len(wb.sheetnames) >= 3:
            print(f"PASS: Component 1 — New sheet '{pivot_sheet_name}' found ({len(wb.sheetnames)} sheets total) (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — No new sheet found. Sheets: {wb.sheetnames}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Early exit if no pivot sheet
    if pivot_sheet_name is None:
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws = wb[pivot_sheet_name]

    # Component 2: Pivot structure — Products as rows, Regions as columns (0.25 points)
    try:
        # Scan for header row containing region names
        header_row = None
        product_col = None
        region_cols = {}

        for r in range(1, min(ws.max_row + 1, 20)):
            row_vals = []
            for c in range(1, min(ws.max_column + 1, 20)):
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    row_vals.append((c, str(val).strip()))

            # Check if this row contains region headers
            found_regions = set()
            for col_idx, val in row_vals:
                for reg in EXPECTED_REGIONS:
                    if reg.lower() == val.lower():
                        found_regions.add(reg)
                        region_cols[reg] = col_idx

            if len(found_regions) >= 3:
                header_row = r
                # Identify product column (usually first non-region column)
                for col_idx, val in row_vals:
                    if val.lower() not in {reg.lower() for reg in EXPECTED_REGIONS} and val.lower() not in ('grand total', 'total'):
                        product_col = col_idx
                        break
                break

        # Scan for product names in the product column
        found_products = set()
        if product_col and header_row:
            for r in range(header_row + 1, min(ws.max_row + 1, 50)):
                val = ws.cell(row=r, column=product_col).value
                if val is not None:
                    val_str = str(val).strip()
                    for prod in EXPECTED_PRODUCTS:
                        if prod.lower() == val_str.lower():
                            found_products.add(prod)

        has_regions = len(region_cols) >= 3
        has_products = len(found_products) >= 3

        if has_regions and has_products:
            print(f"PASS: Component 2 — Pivot structure correct. Products: {found_products}, Region cols: {list(region_cols.keys())} (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — Incomplete pivot structure. Regions found: {list(region_cols.keys())}, Products found: {found_products}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")
        header_row = None
        product_col = None
        region_cols = {}
        found_products = set()

    # Component 3: Revenue data values match source sums (0.25 points)
    try:
        source_totals = compute_source_totals(wb)

        if not header_row or not product_col or not region_cols:
            print(f"FAIL: Component 3 — Cannot verify data without valid pivot structure")
        else:
            # Build map of product rows
            product_rows = {}
            for r in range(header_row + 1, min(ws.max_row + 1, 50)):
                val = ws.cell(row=r, column=product_col).value
                if val is not None:
                    val_str = str(val).strip()
                    for prod in EXPECTED_PRODUCTS:
                        if prod.lower() == val_str.lower():
                            product_rows[prod] = r

            correct_cells = 0
            total_cells = 0
            for prod in EXPECTED_PRODUCTS:
                if prod not in product_rows:
                    continue
                r = product_rows[prod]
                for reg in EXPECTED_REGIONS:
                    if reg not in region_cols:
                        continue
                    total_cells += 1
                    cell_val = ws.cell(row=r, column=region_cols[reg]).value
                    expected_val = source_totals.get(prod, {}).get(reg, 0)
                    if cell_val is not None:
                        try:
                            if abs(float(cell_val) - expected_val) < 1.0:
                                correct_cells += 1
                            else:
                                print(f"  DATA MISMATCH: {prod}/{reg}: expected ~{expected_val:.2f}, found {cell_val}")
                        except (ValueError, TypeError):
                            print(f"  DATA ERROR: {prod}/{reg}: non-numeric value {cell_val}")
                    else:
                        print(f"  DATA MISSING: {prod}/{reg}: cell is empty")

            if total_cells > 0 and correct_cells / total_cells >= 0.8:
                print(f"PASS: Component 3 — {correct_cells}/{total_cells} data cells match source totals (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 3 — Only {correct_cells}/{total_cells} data cells match source totals")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Grand total is approximately $844,316.62 (0.25 points)
    try:
        # Search for grand total value in the pivot sheet
        grand_total_cell = None
        for r in range(1, min(ws.max_row + 1, 50)):
            for c in range(1, min(ws.max_column + 1, 20)):
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    try:
                        num_val = float(val)
                        if abs(num_val - EXPECTED_GRAND_TOTAL) < 100.0:
                            grand_total_cell = (r, c, num_val)
                    except (ValueError, TypeError):
                        pass

        if grand_total_cell is not None:
            r, c, num_val = grand_total_cell
            print(f"PASS: Component 4 — Grand total {num_val:.2f} found at {ws.cell(row=r, column=c).coordinate} (expected ~{EXPECTED_GRAND_TOTAL:.2f}) (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 4 — Grand total ~{EXPECTED_GRAND_TOTAL:.2f} not found in pivot sheet")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
