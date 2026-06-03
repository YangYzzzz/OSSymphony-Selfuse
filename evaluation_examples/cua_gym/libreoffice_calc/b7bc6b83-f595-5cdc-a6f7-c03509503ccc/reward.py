"""
Reward Script: Add medium-weight bottom border to header row cells A1:G1 only
Task ID: calc_fmt_border_bottom_only_014
Domain: libreoffice_calc
Scoring:
  Component 1 (0.6): All 7 cells A1:G1 have medium bottom border
  Component 2 (0.3): All 7 cells A1:G1 have medium bottom border AND no top/left/right borders
                     (compound check: verifies the task was done correctly/precisely)
  Component 3 (0.1): Header cell values remain unchanged AND bottom border is present
                     (compound integrity check, gated on task completion)
Total: 1.0
"""

import os
import openpyxl
from openpyxl.utils import get_column_letter

WORKDIR = '/home/user'  # VM path — reward scripts run on the VM
TASK_ID = 'calc_fmt_border_bottom_only_014'
SHEET_NAME = 'HR Records'
HEADER_COLS = 7  # A through G (columns 1-7)

EXPECTED_HEADERS = [
    'Employee ID', 'First Name', 'Last Name',
    'Department', 'Hire Date', 'Salary', 'Manager'
]


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: sheet must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Gather border info for all header cells once
    border_info = {}
    for col in range(1, HEADER_COLS + 1):
        cell = ws.cell(row=1, column=col)
        border = cell.border
        border_info[col] = {
            'bottom': border.bottom.style if border.bottom else None,
            'top': border.top.style if border.top else None,
            'left': border.left.style if border.left else None,
            'right': border.right.style if border.right else None,
        }

    # Component 1: All 7 header cells A1:G1 have medium bottom border (0.6 points)
    # FAILS on initial (no borders exist) — PASSES on golden (medium bottom present)
    try:
        cells_with_medium_bottom = 0
        for col in range(1, HEADER_COLS + 1):
            if border_info[col]['bottom'] == 'medium':
                cells_with_medium_bottom += 1
            else:
                print(f"  FAIL: {get_column_letter(col)}1 bottom border is "
                      f"'{border_info[col]['bottom']}', expected 'medium'")

        if cells_with_medium_bottom == HEADER_COLS:
            print(f"PASS: Component 1 — All {HEADER_COLS} header cells (A1:G1) have "
                  f"medium bottom border (0.6 pts)")
            total_score += 0.6
        else:
            print(f"FAIL: Component 1 — Only {cells_with_medium_bottom}/{HEADER_COLS} "
                  f"header cells have medium bottom border")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All 7 header cells have medium bottom border AND no top/left/right borders (0.3 points)
    # This is a compound check: verifies correctness/precision of the task (bottom-only).
    # FAILS on initial (no medium bottom present at all) — PASSES on golden (bottom=medium, others=None)
    # If agent adds extra borders, this also fails appropriately.
    try:
        extra_borders_found = []

        # Sub-check A: all cells must have medium bottom (same as comp1 — acts as gate)
        all_medium_bottom = (cells_with_medium_bottom == HEADER_COLS)
        if not all_medium_bottom:
            print(f"  FAIL: Component 2 sub-check A — not all cells have medium bottom border")

        # Sub-check B: no top/left/right borders on any of the 7 cells
        for col in range(1, HEADER_COLS + 1):
            top = border_info[col]['top']
            left = border_info[col]['left']
            right = border_info[col]['right']
            if top is not None:
                extra_borders_found.append(f"{get_column_letter(col)}1 has unexpected top={top}")
            if left is not None:
                extra_borders_found.append(f"{get_column_letter(col)}1 has unexpected left={left}")
            if right is not None:
                extra_borders_found.append(f"{get_column_letter(col)}1 has unexpected right={right}")

        if all_medium_bottom and not extra_borders_found:
            print(f"PASS: Component 2 — All A1:G1 cells have medium bottom border "
                  f"and no other borders (0.3 pts)")
            total_score += 0.3
        else:
            if extra_borders_found:
                print(f"FAIL: Component 2 — Unexpected extra borders found:")
                for msg in extra_borders_found:
                    print(f"  {msg}")
            else:
                print(f"FAIL: Component 2 — Compound check failed (see Component 1 failures)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Cell values in row 1 remain unchanged AND medium bottom border is present (0.1 points)
    # Compound check: gated on bottom border having been applied (to avoid scoring pre-existing values).
    # FAILS on initial (no medium bottom border) — PASSES on golden (border present + values intact)
    try:
        has_all_medium_bottom = (cells_with_medium_bottom == HEADER_COLS)

        if not has_all_medium_bottom:
            print(f"FAIL: Component 3 — Skipped (medium bottom border not applied; "
                  f"precondition for this check not met)")
        else:
            mismatched_headers = []
            for col, expected_header in enumerate(EXPECTED_HEADERS, 1):
                actual = ws.cell(row=1, column=col).value
                if actual != expected_header:
                    print(f"  FAIL: {get_column_letter(col)}1 = {repr(actual)}, "
                          f"expected {repr(expected_header)}")
                    mismatched_headers.append(col)

            if not mismatched_headers:
                print(f"PASS: Component 3 — All header cell values remain unchanged "
                      f"(verified with border present) (0.1 pts)")
                total_score += 0.1
            else:
                print(f"FAIL: Component 3 — Header values were modified")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
