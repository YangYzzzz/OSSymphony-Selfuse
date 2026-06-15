"""
Initial Setup: HR Records spreadsheet with no borders applied
Task ID: calc_fmt_border_bottom_only_014
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font
import datetime

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_border_bottom_only_014'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: HR Records ---
    ws = wb.active
    ws.title = 'HR Records'

    # Headers (row 1) — NO borders applied
    headers = ['Employee ID', 'First Name', 'Last Name', 'Department', 'Hire Date', 'Salary', 'Manager']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic employee data (rows 2–50, 49 employees)
    departments = ['Engineering', 'Marketing', 'Finance', 'HR', 'Operations', 'Sales', 'Legal', 'IT Support']
    managers = [
        'David Kim', 'Rachel Torres', 'Steven Brooks', 'Angela White',
        'Frank Chen', 'Linda Park', 'Michael Scott', 'Nina Patel'
    ]

    employees = [
        ('EMP001', 'Sarah',     'Chen',       'Engineering',  datetime.date(2019, 3, 15),  95000, 'David Kim'),
        ('EMP002', 'Marcus',    'Johnson',     'Marketing',    datetime.date(2020, 6, 1),   72000, 'Rachel Torres'),
        ('EMP003', 'Priya',     'Sharma',      'Finance',      datetime.date(2018, 11, 20), 88000, 'Steven Brooks'),
        ('EMP004', 'Carlos',    'Mendoza',     'HR',           datetime.date(2021, 2, 8),   65000, 'Angela White'),
        ('EMP005', 'Emma',      'Williams',    'Engineering',  datetime.date(2017, 9, 5),  102000, 'David Kim'),
        ('EMP006', 'James',     'Okafor',      'Operations',   datetime.date(2022, 4, 18),  61000, 'Frank Chen'),
        ('EMP007', 'Nina',      'Patel',       'IT Support',   datetime.date(2016, 7, 22),  78000, 'Michael Scott'),
        ('EMP008', 'Thomas',    'Lee',         'Sales',        datetime.date(2020, 10, 3),  69500, 'Linda Park'),
        ('EMP009', 'Olivia',    'Nguyen',      'Engineering',  datetime.date(2023, 1, 9),   91000, 'David Kim'),
        ('EMP010', 'Samuel',    'Brooks',      'Finance',      datetime.date(2019, 5, 14),  84000, 'Steven Brooks'),
        ('EMP011', 'Fatima',    'Al-Hassan',   'Legal',        datetime.date(2015, 8, 30),  97000, 'Angela White'),
        ('EMP012', 'Lucas',     'Andersen',    'Marketing',    datetime.date(2021, 11, 25), 74000, 'Rachel Torres'),
        ('EMP013', 'Yuki',      'Tanaka',      'Engineering',  datetime.date(2018, 3, 12),  99000, 'David Kim'),
        ('EMP014', 'Amara',     'Diallo',      'HR',           datetime.date(2022, 6, 7),   63000, 'Angela White'),
        ('EMP015', 'Ethan',     'Murphy',      'Operations',   datetime.date(2020, 9, 19),  58000, 'Frank Chen'),
        ('EMP016', 'Sofia',     'Reyes',       'Sales',        datetime.date(2019, 12, 4),  71000, 'Linda Park'),
        ('EMP017', 'Benjamin',  'Clarke',      'Finance',      datetime.date(2017, 4, 27),  86000, 'Steven Brooks'),
        ('EMP018', 'Hana',      'Yamamoto',    'IT Support',   datetime.date(2023, 3, 3),   76000, 'Michael Scott'),
        ('EMP019', 'Anthony',   'Brown',       'Legal',        datetime.date(2014, 10, 16), 105000, 'Angela White'),
        ('EMP020', 'Isabella',  'Ferreira',    'Engineering',  datetime.date(2021, 7, 21),  93000, 'David Kim'),
        ('EMP021', 'Daniel',    'Kim',         'Marketing',    datetime.date(2022, 2, 14),  68000, 'Rachel Torres'),
        ('EMP022', 'Aisha',     'Mohammed',    'HR',           datetime.date(2018, 6, 9),   67000, 'Angela White'),
        ('EMP023', 'Noah',      'Harrison',    'Operations',   datetime.date(2020, 1, 28),  60000, 'Frank Chen'),
        ('EMP024', 'Claire',    'Beaumont',    'Finance',      datetime.date(2016, 5, 11),  89000, 'Steven Brooks'),
        ('EMP025', 'Raj',       'Krishnamurthy','IT Support',  datetime.date(2019, 8, 23),  81000, 'Michael Scott'),
        ('EMP026', 'Mia',       'Johansson',   'Sales',        datetime.date(2023, 5, 6),   66000, 'Linda Park'),
        ('EMP027', 'Liam',      'O\'Brien',    'Engineering',  datetime.date(2015, 2, 17),  110000, 'David Kim'),
        ('EMP028', 'Grace',     'Liu',         'Legal',        datetime.date(2020, 11, 29), 94000, 'Angela White'),
        ('EMP029', 'Aaron',     'Goldstein',   'Marketing',    datetime.date(2017, 7, 4),   77000, 'Rachel Torres'),
        ('EMP030', 'Zara',      'Hussain',     'Finance',      datetime.date(2022, 9, 15),  82000, 'Steven Brooks'),
        ('EMP031', 'Elijah',    'Watts',       'HR',           datetime.date(2021, 3, 22),  64000, 'Angela White'),
        ('EMP032', 'Nadia',     'Petrov',      'Engineering',  datetime.date(2018, 12, 10), 96000, 'David Kim'),
        ('EMP033', 'Victor',    'Almeida',     'Operations',   datetime.date(2019, 4, 7),   59500, 'Frank Chen'),
        ('EMP034', 'Hannah',    'Schultz',     'Sales',        datetime.date(2022, 1, 18),  70000, 'Linda Park'),
        ('EMP035', 'Kwame',     'Asante',      'IT Support',   datetime.date(2016, 10, 25), 80000, 'Michael Scott'),
        ('EMP036', 'Lily',      'Chen',        'Legal',        datetime.date(2023, 7, 12),   91500, 'Angela White'),
        ('EMP037', 'Omar',      'Hassan',      'Finance',      datetime.date(2020, 3, 31),  85000, 'Steven Brooks'),
        ('EMP038', 'Camila',    'Gutierrez',   'Marketing',    datetime.date(2017, 11, 16), 73000, 'Rachel Torres'),
        ('EMP039', 'Dominic',   'Russo',       'Engineering',  datetime.date(2021, 8, 3),   98000, 'David Kim'),
        ('EMP040', 'Leila',     'Nazari',      'HR',           datetime.date(2019, 6, 20),  66500, 'Angela White'),
        ('EMP041', 'Patrick',   'Fitzgerald',  'Operations',   datetime.date(2015, 4, 9),   62000, 'Frank Chen'),
        ('EMP042', 'Selin',     'Yilmaz',      'Sales',        datetime.date(2022, 10, 27), 68500, 'Linda Park'),
        ('EMP043', 'Andre',     'Dupont',      'IT Support',   datetime.date(2018, 1, 14),  79500, 'Michael Scott'),
        ('EMP044', 'Bianca',    'Santos',      'Engineering',  datetime.date(2020, 5, 25),  92000, 'David Kim'),
        ('EMP045', 'Miles',     'Thompson',    'Finance',      datetime.date(2016, 9, 8),   87000, 'Steven Brooks'),
        ('EMP046', 'Freya',     'Larsen',      'Legal',        datetime.date(2023, 2, 21),  93500, 'Angela White'),
        ('EMP047', 'Jerome',    'Washington',  'Marketing',    datetime.date(2019, 10, 12), 75000, 'Rachel Torres'),
        ('EMP048', 'Tara',      'Benson',      'HR',           datetime.date(2021, 6, 17),  63500, 'Angela White'),
        ('EMP049', 'Diego',     'Morales',     'Engineering',  datetime.date(2017, 1, 30),  100000, 'David Kim'),
    ]

    for r, emp in enumerate(employees, 2):
        for c, val in enumerate(emp, 1):
            ws.cell(row=r, column=c, value=val)

    # Column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

create_initial()
