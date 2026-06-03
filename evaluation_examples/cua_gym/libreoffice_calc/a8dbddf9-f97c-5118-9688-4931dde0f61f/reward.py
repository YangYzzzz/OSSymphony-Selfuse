"""
Reward Script: Extract NIH R01 grant funding data from PDFs and build a table in LibreOffice Calc
Task ID: osworld_multi_apps_pdf_stats_table_003
Domain: libreoffice_calc
Scoring:
  Component 1: Correct headers (Year, #Submitted, #Funded, Funding Rate (%)) — 0.3 pts
  Component 2: All 5 data rows present for years 2018-2022 with submitted/funded counts — 0.3 pts
  Component 3: Funding rates correctly computed and formatted to 2 decimal places — 0.4 pts
Total: 1.0
"""

import os

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_pdf_stats_table_003'
FILE_PATH = f'{WORKDIR}/Desktop/NIH_R01_rates.xlsx'

# Expected ground truth data (from task context)
EXPECTED_HEADERS = ['Year', '#Submitted', '#Funded', 'Funding Rate (%)']
EXPECTED_DATA = {
    2018: {'submitted': 4631, 'funded': 850, 'rate': 18.35},
    2019: {'submitted': 4892, 'funded': 891, 'rate': 18.21},
    2020: {'submitted': 5124, 'funded': 942, 'rate': 18.38},
    2021: {'submitted': 5380, 'funded': 991, 'rate': 18.42},
    2022: {'submitted': 5612, 'funded': 1018, 'rate': 18.14},
}
RATE_TOLERANCE = 0.01  # tolerance for floating point comparison


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition gate: file must exist
    if not os.path.exists(file_path):
        print(f"CRITICAL: File not found at {file_path}")
        print("REWARD: 0.0")
        return 0.0

    # Load the workbook
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load workbook {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Find the worksheet (use the first sheet regardless of name)
    ws = wb.worksheets[0]
    print(f"INFO: Using sheet '{ws.title}' with {ws.max_row} rows and {ws.max_column} columns")

    # Read all rows into memory (values only)
    all_rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=4, values_only=True):
        all_rows.append(list(row))

    # -------------------------------------------------------------------------
    # Component 1: Correct headers (0.3 points)
    # Headers must be: Year, #Submitted, #Funded, Funding Rate (%)
    # This FAILS on initial_env (no file) → PASSES on golden_env
    # -------------------------------------------------------------------------
    try:
        if len(all_rows) < 1:
            print("FAIL: Component 1 — No rows found in worksheet")
        else:
            header_row = all_rows[0]
            # Normalize for comparison (strip whitespace, case-insensitive for robustness)
            found_headers = [str(h).strip() if h is not None else '' for h in header_row]
            expected_lower = [h.lower() for h in EXPECTED_HEADERS]
            found_lower = [h.lower() for h in found_headers]

            headers_match = (found_lower == expected_lower)
            if headers_match:
                print(f"PASS: Component 1 — Correct headers found: {found_headers} (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 1 — Expected headers {EXPECTED_HEADERS}, found {found_headers}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: All 5 data rows for years 2018-2022 with correct submitted/funded (0.3 points)
    # This FAILS on initial_env (no file) → PASSES on golden_env
    # -------------------------------------------------------------------------
    try:
        if len(all_rows) < 2:
            print("FAIL: Component 2 — No data rows found (only header or empty file)")
        else:
            data_rows = all_rows[1:]  # skip header

            # Build a dict from year → row
            year_to_row = {}
            for row in data_rows:
                if row[0] is not None:
                    try:
                        year_val = int(row[0])
                        year_to_row[year_val] = row
                    except (ValueError, TypeError):
                        pass

            missing_years = []
            wrong_counts = []
            for year, expected in EXPECTED_DATA.items():
                if year not in year_to_row:
                    missing_years.append(year)
                    continue
                row = year_to_row[year]
                try:
                    submitted = int(row[1]) if row[1] is not None else None
                    funded = int(row[2]) if row[2] is not None else None
                except (ValueError, TypeError):
                    submitted = None
                    funded = None

                if submitted != expected['submitted'] or funded != expected['funded']:
                    wrong_counts.append(
                        f"Year {year}: expected submitted={expected['submitted']}, funded={expected['funded']}; "
                        f"got submitted={submitted}, funded={funded}"
                    )

            if not missing_years and not wrong_counts:
                print(f"PASS: Component 2 — All 5 data rows (2018-2022) with correct submitted/funded counts (0.3 pts)")
                total_score += 0.3
            else:
                if missing_years:
                    print(f"FAIL: Component 2 — Missing years: {missing_years}")
                if wrong_counts:
                    for msg in wrong_counts:
                        print(f"FAIL: Component 2 — {msg}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Funding rates correctly computed and formatted to 2 decimal places (0.4 points)
    # Funding Rate (%) = (#Funded / #Submitted) * 100, rounded to 2 decimal places
    # This FAILS on initial_env (no file) → PASSES on golden_env
    # -------------------------------------------------------------------------
    try:
        if len(all_rows) < 2:
            print("FAIL: Component 3 — No data rows found")
        else:
            data_rows = all_rows[1:]

            # Build year → row dict again (or reuse)
            year_to_row_2 = {}
            for row in data_rows:
                if row[0] is not None:
                    try:
                        year_val = int(row[0])
                        year_to_row_2[year_val] = row
                    except (ValueError, TypeError):
                        pass

            rate_errors = []
            rate_passes = 0
            for year, expected in EXPECTED_DATA.items():
                if year not in year_to_row_2:
                    rate_errors.append(f"Year {year}: row missing")
                    continue
                row = year_to_row_2[year]
                rate_val = row[3]
                if rate_val is None:
                    rate_errors.append(f"Year {year}: funding rate is None")
                    continue
                try:
                    rate_float = float(rate_val)
                except (ValueError, TypeError):
                    rate_errors.append(f"Year {year}: funding rate not numeric: {rate_val}")
                    continue

                expected_rate = expected['rate']
                if abs(rate_float - expected_rate) <= RATE_TOLERANCE:
                    rate_passes += 1
                else:
                    # Also allow rate computed from actual submitted/funded values
                    try:
                        submitted = int(row[1])
                        funded = int(row[2])
                        computed_rate = round((funded / submitted) * 100, 2)
                        if abs(rate_float - computed_rate) <= RATE_TOLERANCE:
                            rate_passes += 1
                        else:
                            rate_errors.append(
                                f"Year {year}: expected rate ~{expected_rate}, found {rate_float}"
                            )
                    except Exception:
                        rate_errors.append(
                            f"Year {year}: expected rate ~{expected_rate}, found {rate_float}"
                        )

            if rate_passes == 5 and not rate_errors:
                print(f"PASS: Component 3 — All 5 funding rates correctly computed to 2 decimal places (0.4 pts)")
                total_score += 0.4
            elif rate_passes > 0:
                print(f"FAIL: Component 3 — only {rate_passes}/5 rates correct")
                for err in rate_errors:
                    print(f"  FAIL: {err}")
                # Award partial credit: 0.08 per correct rate
                partial = round(rate_passes * 0.08, 2)
                print(f"PARTIAL: Component 3 — Awarding {partial} pts for {rate_passes}/5 correct rates")
                if rate_passes > 0:
                    total_score += partial
            else:
                print(f"FAIL: Component 3 — 0/5 rates correct")
                for err in rate_errors:
                    print(f"  FAIL: {err}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Run verification against the canonical artifact path
if not os.path.exists(FILE_PATH):
    print(f"File not found: {FILE_PATH}")
    print("REWARD: 0.0")
else:
    verify_task(FILE_PATH)
