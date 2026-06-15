"""
Reward Script: Extract first authors from arXiv PDFs and create spreadsheet
Task ID: osworld_multi_apps_pdf_author_extract_009
Domain: libreoffice_calc (multi-apps: PDF reading + spreadsheet creation)
Scoring:
  - Component 1: Correct headers (Name, Email, Affiliation, Year)          — 0.20 pts
  - Component 2: Correct row count (9 data rows)                           — 0.10 pts
  - Component 3: Correct author data for all 9 rows (Name/Email/Affil/Year)— 0.45 pts
  - Component 4: Correct sort order (Year desc, then Name asc)             — 0.25 pts
  Total: 1.0
"""

import os

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_pdf_author_extract_009'
FILE_PATH = f'{WORKDIR}/preprint_authors.xlsx'

# Ground truth: 9 rows extracted from 9 arXiv PDF preprints
# Sorted by Year descending, then by Name ascending within same year
EXPECTED_DATA = [
    # (Name, Email, Affiliation, Year)
    ('Baptiste Roziere', 'roziere@meta.com', 'Meta AI', 2024),
    ('Nisan Stiennon', 'nisan@openai.com', 'OpenAI', 2024),
    ('Amanda Askell', 'amanda@anthropic.com', 'Anthropic', 2023),
    ('Haotian Liu', 'haotianliu@cs.wisc.edu', 'University of Wisconsin-Madison', 2023),
    ('Patrick Lewis', 'plewis@meta.com', 'Meta AI Research', 2023),
    ('Jared Kaplan', 'jkaplan@openai.com', 'OpenAI', 2022),
    ('Zhen Yang', 'zhen.yang@bytedance.com', 'ByteDance Research', 2022),
    ('Jonathan Ho', 'jonathanho@google.com', 'Google Brain', 2021),
    ('Ting Chen', 'tingchen@google.com', 'Google Research, Brain Team', 2021),
]

EXPECTED_HEADERS = ['Name', 'Email', 'Affiliation', 'Year']


def normalize(val):
    """Normalize a cell value for comparison."""
    if val is None:
        return ''
    return str(val).strip()


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: openpyxl must be available
    if not OPENPYXL_AVAILABLE:
        print("CRITICAL: openpyxl not available — cannot verify")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: file must exist and be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get the active/first sheet
    try:
        ws = wb.active
    except Exception as e:
        print(f"CRITICAL: Cannot access worksheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Correct headers — 'Name', 'Email', 'Affiliation', 'Year' (0.20 points)
    try:
        actual_headers = [normalize(ws.cell(1, c).value) for c in range(1, 5)]
        expected_headers_norm = [h.strip() for h in EXPECTED_HEADERS]
        if actual_headers == expected_headers_norm:
            print(f"PASS: Component 1 — Headers correct: {actual_headers} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 — Expected headers {EXPECTED_HEADERS}, found {actual_headers}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Correct row count — exactly 9 data rows (0.10 points)
    try:
        # Count non-empty rows starting from row 2
        data_rows = 0
        for row_idx in range(2, ws.max_row + 1):
            row_vals = [ws.cell(row_idx, c).value for c in range(1, 5)]
            if any(v is not None for v in row_vals):
                data_rows += 1
        if data_rows == 9:
            print(f"PASS: Component 2 — Row count correct: {data_rows} data rows (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 — Expected 9 data rows, found {data_rows}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Correct author data for all 9 rows (0.45 points)
    # Award 0.05 per row that has all 4 fields correct
    try:
        matched_rows = 0
        for i, (exp_name, exp_email, exp_affil, exp_year) in enumerate(EXPECTED_DATA):
            row_idx = i + 2  # row 1 is header
            act_name = normalize(ws.cell(row_idx, 1).value)
            act_email = normalize(ws.cell(row_idx, 2).value)
            act_affil = normalize(ws.cell(row_idx, 3).value)
            act_year_raw = ws.cell(row_idx, 4).value
            try:
                act_year = int(act_year_raw) if act_year_raw is not None else None
            except (ValueError, TypeError):
                act_year = None

            name_ok = act_name == exp_name.strip()
            email_ok = act_email == exp_email.strip()
            affil_ok = act_affil == exp_affil.strip()
            year_ok = act_year == exp_year

            if name_ok and email_ok and affil_ok and year_ok:
                matched_rows += 1
                print(f"  PASS: Row {row_idx} — {exp_name} ({exp_year})")
            else:
                issues = []
                if not name_ok:
                    issues.append(f"Name: expected '{exp_name}', got '{act_name}'")
                if not email_ok:
                    issues.append(f"Email: expected '{exp_email}', got '{act_email}'")
                if not affil_ok:
                    issues.append(f"Affil: expected '{exp_affil}', got '{act_affil}'")
                if not year_ok:
                    issues.append(f"Year: expected {exp_year}, got {act_year}")
                print(f"  FAIL: Row {row_idx} — {'; '.join(issues)}")

        row_score = round(matched_rows * 0.05, 4)
        if matched_rows == 9:
            print(f"PASS: Component 3 — All 9 rows correct (0.45 pts)")
        else:
            print(f"PARTIAL: Component 3 — {matched_rows}/9 rows correct ({row_score:.2f} pts)")
        total_score += row_score
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Correct sort order — Year descending, then Name ascending (0.25 points)
    # Check that the rows in the spreadsheet follow the expected sorted order
    try:
        actual_years = []
        actual_names = []
        for row_idx in range(2, 11):  # rows 2-10 (9 data rows)
            year_raw = ws.cell(row_idx, 4).value
            name_raw = normalize(ws.cell(row_idx, 1).value)
            try:
                yr = int(year_raw) if year_raw is not None else 0
            except (ValueError, TypeError):
                yr = 0
            actual_years.append(yr)
            actual_names.append(name_raw)

        # Build expected order from EXPECTED_DATA (already sorted correctly)
        expected_years = [r[3] for r in EXPECTED_DATA]
        expected_names = [r[0] for r in EXPECTED_DATA]

        sort_correct = (actual_years == expected_years and actual_names == expected_names)

        if sort_correct:
            print(f"PASS: Component 4 — Sort order correct: Year desc then Name asc (0.25 pts)")
            total_score += 0.25
        else:
            # Check partial: only years are in correct descending order
            years_desc = all(actual_years[i] >= actual_years[i+1] for i in range(len(actual_years)-1))
            if years_desc:
                print(f"PARTIAL: Component 4 — Year ordering correct but Name ordering within year may be wrong")
                print(f"  Expected names: {expected_names}")
                print(f"  Actual names:   {actual_names}")
            else:
                print(f"FAIL: Component 4 — Sort order incorrect")
                print(f"  Expected years: {expected_years}")
                print(f"  Actual years:   {actual_years}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Run verification against canonical artifact path
if not OPENPYXL_AVAILABLE:
    print("openpyxl not available — cannot verify")
    print("REWARD: 0.0")
elif not os.path.exists(FILE_PATH):
    print(f"File not found: {FILE_PATH}")
    print("REWARD: 0.0")
else:
    verify_task(FILE_PATH)
