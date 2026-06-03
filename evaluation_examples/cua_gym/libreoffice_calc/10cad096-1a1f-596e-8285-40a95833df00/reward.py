"""
Reward Script: Pivot table showing MIN and MAX temperature by station and month
Task ID: calc_pivot_068
Domain: libreoffice_calc
Scoring:
  Component 1 (0.15): PivotTable sheet exists
  Component 2 (0.25): Correct structure (headers, stations, 24 data columns)
  Component 3 (0.30): Station_A ground truth values match context
  Component 4 (0.30): All stations present with complete numeric data
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_068'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: PivotTable sheet exists (0.15 points)
    # This should FAIL on initial (only WeatherData) and PASS on golden
    try:
        pivot_sheet_found = False
        for sn in wb.sheetnames:
            if 'pivot' in sn.lower():
                pivot_sheet_found = True
                pivot_sheet_name = sn
                break
        if pivot_sheet_found:
            print(f"PASS: Component 1 -- Pivot sheet found: '{pivot_sheet_name}' (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 -- No sheet containing 'pivot' found. Sheets: {wb.sheetnames}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    if not pivot_sheet_found:
        # No pivot sheet means nothing else to verify
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    ws = wb[pivot_sheet_name]

    # Component 2: Correct structure (0.25 points)
    # Check: header row with Station + 24 month-metric columns, 3 station data rows
    try:
        structure_score = 0.0

        # Find the header row: must have 'Station' as a standalone cell value
        # (not embedded in a title like 'Temperature Summary by Station and Month')
        # AND have month-related columns in the same row
        header_row = None
        for r in range(1, min(ws.max_row + 1, 10)):
            cell_val = ws.cell(r, 1).value
            if cell_val is not None and str(cell_val).strip().lower() == 'station':
                # Verify this row also has month references in other columns
                has_month = False
                for c in range(2, min(ws.max_column + 1, 30)):
                    hv = ws.cell(r, c).value
                    if hv is not None and any(m in str(hv).lower() for m in
                                              ['jan', 'feb', 'mar', 'apr', 'may', 'jun',
                                               'jul', 'aug', 'sep', 'oct', 'nov', 'dec']):
                        has_month = True
                        break
                if has_month:
                    header_row = r
                    break

        if header_row is None:
            print(f"FAIL: Component 2 -- No header row with 'Station' found")
        else:
            # Count data columns that reference months
            month_keywords = ['jan', 'feb', 'mar', 'apr', 'may', 'jun',
                              'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
            month_cols = 0
            for c in range(2, ws.max_column + 1):
                hdr = ws.cell(header_row, c).value
                if hdr is not None:
                    hdr_lower = str(hdr).lower()
                    if any(m in hdr_lower for m in month_keywords):
                        month_cols += 1

            # Check for MIN and MAX in headers
            has_min = False
            has_max = False
            for c in range(2, ws.max_column + 1):
                hdr = ws.cell(header_row, c).value
                if hdr is not None:
                    hdr_lower = str(hdr).lower()
                    if 'min' in hdr_lower:
                        has_min = True
                    if 'max' in hdr_lower:
                        has_max = True

            # Count station data rows (rows below header with station names like Station_A, etc.)
            station_rows = 0
            station_names_found = []
            for r in range(header_row + 1, ws.max_row + 1):
                val = ws.cell(r, 1).value
                if val is not None and 'station_' in str(val).lower().replace(' ', '_'):
                    station_rows += 1
                    station_names_found.append(str(val))

            # Scoring sub-components
            if month_cols >= 20:  # At least 20 month-related columns (expect 24)
                structure_score += 0.10
            if has_min and has_max:
                structure_score += 0.05
            if station_rows >= 3:
                structure_score += 0.10

            if structure_score > 0:
                print(f"PASS: Component 2 -- Structure: {month_cols} month cols, "
                      f"MIN={has_min}, MAX={has_max}, {station_rows} stations: "
                      f"{station_names_found} ({structure_score} pts)")
            else:
                print(f"FAIL: Component 2 -- Structure issues: {month_cols} month cols, "
                      f"MIN={has_min}, MAX={has_max}, {station_rows} stations")
            total_score += structure_score

    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Station_A ground truth values (0.30 points)
    # Expected: Station_A Jan MIN=-3, Jan MAX=12, Jul MIN=18, Jul MAX=36
    try:
        gt_score = 0.0

        # Find Station_A row and identify column mapping from headers
        station_a_row = None
        for r in range(header_row + 1, ws.max_row + 1):
            val = ws.cell(r, 1).value
            if val is not None and 'station_a' in str(val).lower().replace(' ', '_'):
                station_a_row = r
                break

        if station_a_row is None:
            print(f"FAIL: Component 3 -- Station_A row not found")
        else:
            # Build column mapping: header text -> column index
            col_map = {}
            for c in range(2, ws.max_column + 1):
                hdr = ws.cell(header_row, c).value
                if hdr is not None:
                    col_map[str(hdr).strip().lower()] = c

            # Ground truth checks (each worth 0.075 pts)
            checks = [
                ('jan min', -3, 1.0),
                ('jan max', 12, 1.0),
                ('jul min', 18, 1.0),
                ('jul max', 36, 1.0),
            ]

            for label, expected, tol in checks:
                col_idx = None
                for key, idx in col_map.items():
                    if label in key.lower():
                        col_idx = idx
                        break

                if col_idx is None:
                    print(f"  FAIL: Station_A {label} -- column not found in headers")
                    continue

                actual = ws.cell(station_a_row, col_idx).value
                if actual is not None:
                    try:
                        if abs(float(actual) - expected) <= tol:
                            gt_score += 0.075
                            print(f"  PASS: Station_A {label} = {actual} (expected {expected})")
                        else:
                            print(f"  FAIL: Station_A {label} = {actual} (expected {expected})")
                    except (ValueError, TypeError):
                        print(f"  FAIL: Station_A {label} = {actual} (not numeric)")
                else:
                    print(f"  FAIL: Station_A {label} = None")

        if gt_score > 0:
            print(f"PASS: Component 3 -- Station_A ground truth ({gt_score} pts)")
        else:
            print(f"FAIL: Component 3 -- No ground truth values matched")
        total_score += gt_score

    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: All stations present with complete numeric data (0.30 points)
    # Station_B and Station_C rows exist with numeric temperature values
    try:
        station_score = 0.0
        expected_stations = ['station_b', 'station_c']

        for station_label in expected_stations:
            station_row = None
            for r in range(header_row + 1, ws.max_row + 1):
                val = ws.cell(r, 1).value
                if val is not None and station_label in str(val).lower().replace(' ', '_'):
                    station_row = r
                    break

            if station_row is None:
                print(f"  FAIL: {station_label} row not found")
                continue

            # Check that at least 20 of the data columns have numeric values
            numeric_count = 0
            for c in range(2, ws.max_column + 1):
                cell_val = ws.cell(station_row, c).value
                if cell_val is not None:
                    try:
                        float(cell_val)
                        numeric_count += 1
                    except (ValueError, TypeError):
                        pass

            if numeric_count >= 20:
                station_score += 0.15
                print(f"  PASS: {station_label} has {numeric_count} numeric values")
            else:
                print(f"  FAIL: {station_label} has only {numeric_count} numeric values (need >= 20)")

        if station_score > 0:
            print(f"PASS: Component 4 -- Station data completeness ({station_score} pts)")
        else:
            print(f"FAIL: Component 4 -- Missing station data")
        total_score += station_score

    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
