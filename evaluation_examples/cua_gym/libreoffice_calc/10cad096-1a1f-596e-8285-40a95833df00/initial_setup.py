"""
Initial Setup: Weather data spreadsheet for pivot table task
Task ID: calc_pivot_068
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random
from datetime import datetime, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_068'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    try:
        import openpyxl
    except ImportError:
        subprocess.check_call(['pip3', 'install', 'openpyxl'], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        import openpyxl

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'WeatherData'

    # Headers
    headers = ['ReadingID', 'Date', 'Station', 'Temperature', 'Humidity', 'WindSpeed']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12

    stations = ['Station_A', 'Station_B', 'Station_C']

    # Temperature ranges by month (index 0=Jan, 11=Dec) - min_low, min_high, max_low, max_high
    # These define the range from which we sample min and max temps
    temp_profiles = {
        'Station_A': [
            (-5, 5, 5, 12),    # Jan
            (-3, 6, 6, 14),    # Feb
            (0, 8, 10, 18),    # Mar
            (5, 12, 15, 22),   # Apr
            (10, 16, 20, 28),  # May
            (14, 20, 25, 33),  # Jun
            (18, 24, 28, 36),  # Jul
            (17, 23, 27, 35),  # Aug
            (12, 18, 22, 30),  # Sep
            (6, 12, 14, 22),   # Oct
            (0, 7, 8, 16),     # Nov
            (-4, 4, 4, 12),    # Dec
        ],
        'Station_B': [
            (-2, 4, 6, 13),    # Jan
            (-1, 5, 7, 15),    # Feb
            (2, 9, 11, 19),    # Mar
            (6, 13, 16, 23),   # Apr
            (11, 17, 21, 29),  # May
            (15, 21, 26, 34),  # Jun
            (19, 25, 29, 37),  # Jul
            (18, 24, 28, 36),  # Aug
            (13, 19, 23, 31),  # Sep
            (7, 13, 15, 23),   # Oct
            (1, 8, 9, 17),     # Nov
            (-3, 5, 5, 13),    # Dec
        ],
        'Station_C': [
            (-3, 3, 4, 11),    # Jan
            (-2, 4, 5, 13),    # Feb
            (1, 7, 9, 17),     # Mar
            (4, 11, 14, 21),   # Apr
            (9, 15, 19, 27),   # May
            (13, 19, 24, 32),  # Jun
            (17, 23, 27, 35),  # Jul
            (16, 22, 26, 34),  # Aug
            (11, 17, 21, 29),  # Sep
            (5, 11, 13, 21),   # Oct
            (-1, 6, 7, 15),    # Nov
            (-4, 3, 3, 11),    # Dec
        ],
    }

    # Generate 730 readings: twice daily for 365 days across 3 stations
    # We cycle through stations to distribute evenly
    # 365 days * 2 readings/day = 730, but we need to distribute across 3 stations
    # Actually: 730 rows total, ~243 per station, ~20 per station per month

    start_date = datetime(2024, 1, 1)
    reading_id = 1
    row = 2

    # We need to ensure specific ground truth values:
    # Station_A Jan MIN=-3, MAX=12, Station_A Jul MIN=18, MAX=36
    # We'll track and force these values

    forced_values = {}  # (station, month) -> list of forced temps

    # Force Station_A Jan to have exactly -3 and 12
    forced_values[('Station_A', 1)] = [-3, 12]
    # Force Station_A Jul to have exactly 18 and 36
    forced_values[('Station_A', 7)] = [18, 36]

    all_readings = []

    for day_offset in range(365):
        current_date = start_date + timedelta(days=day_offset)
        month_idx = current_date.month - 1

        # Two readings per day, cycling through stations
        for reading_num in range(2):
            station = stations[(day_offset * 2 + reading_num) % 3]
            profile = temp_profiles[station][month_idx]
            min_low, min_high, max_low, max_high = profile

            # Generate temperature within range
            temp = round(random.uniform(min_low, max_high), 1)

            # Humidity 30-95%
            humidity = round(random.uniform(30, 95), 1)

            # Wind speed 0-45 km/h
            wind_speed = round(random.uniform(0, 45), 1)

            all_readings.append({
                'id': reading_id,
                'date': current_date.strftime('%m/%d/%Y'),
                'station': station,
                'temp': temp,
                'humidity': humidity,
                'wind_speed': wind_speed,
                'month': current_date.month,
            })
            reading_id += 1

    # Now enforce ground truth values
    # For Station_A Jan: ensure min is exactly -3 and max is exactly 12
    for key, forced_temps in forced_values.items():
        station_name, month = key
        matching = [r for r in all_readings if r['station'] == station_name and r['month'] == month]
        if len(matching) >= 2:
            # Set one reading to the forced min, another to forced max
            matching[0]['temp'] = forced_temps[0]
            matching[1]['temp'] = forced_temps[1]
            # Clamp all other readings for this station/month between forced min and max
            for r in matching[2:]:
                r['temp'] = max(forced_temps[0], min(forced_temps[1], r['temp']))

    # Write all readings
    for r_data in all_readings:
        ws.cell(row=row, column=1, value=r_data['id'])
        ws.cell(row=row, column=2, value=r_data['date'])
        ws.cell(row=row, column=3, value=r_data['station'])
        ws.cell(row=row, column=4, value=r_data['temp'])
        ws.cell(row=row, column=5, value=r_data['humidity'])
        ws.cell(row=row, column=6, value=r_data['wind_speed'])
        row += 1

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Total readings: {len(all_readings)}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
