"""
Initial Setup: Add 4 coffee shops to existing Calc file 'coffee_spots.ods'
Task ID: osworld_multi_apps_web_location_002
Domain: libreoffice_calc

Creates coffee_spots.ods on the Desktop with 3 existing rows.
The agent must append 4 new coffee shop rows.
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_web_location_002'
DESKTOP = f'{WORKDIR}/Desktop'
# Final output as .ods on Desktop
OUTPUT_ODS = f'{DESKTOP}/coffee_spots.ods'
# Intermediate xlsx for conversion
OUTPUT_XLSX = f'{DESKTOP}/coffee_spots.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    # Ensure Desktop directory exists
    os.makedirs(DESKTOP, exist_ok=True)

    # Build the initial workbook with 3 realistic coffee shop rows
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Coffee Spots'

    # Headers
    headers = ['Name', 'City', 'Country', 'Rating', 'Specialty', 'Address']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 3 existing rows — realistic coffee shop data (NOT the 4 to be added)
    data = [
        ['Verve Coffee Roasters', 'Santa Cruz', 'USA', 4.7, 'Specialty espresso', '104 Bronson St, Santa Cruz'],
        ['Four Barrel Coffee',   'San Francisco', 'USA', 4.2, 'Seasonal roasts',   '375 Valencia St, SF'],
        ['Ritual Coffee',        'San Francisco', 'USA', 4.5, 'Single-origin',     '1026 Valencia St, SF'],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Save as xlsx first
    wb.save(OUTPUT_XLSX)
    print(f'Intermediate xlsx created: {OUTPUT_XLSX}')

    # Convert xlsx → ods using LibreOffice headless
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    result = subprocess.run(
        [
            'libreoffice', '--headless', '--convert-to', 'ods',
            '--outdir', DESKTOP,
            OUTPUT_XLSX,
        ],
        capture_output=True,
        text=True,
        env=env,
        timeout=60,
    )
    print('LibreOffice convert stdout:', result.stdout)
    print('LibreOffice convert stderr:', result.stderr)

    # Remove the intermediate xlsx
    if os.path.exists(OUTPUT_XLSX):
        os.remove(OUTPUT_XLSX)
        print(f'Removed intermediate file: {OUTPUT_XLSX}')

    if os.path.exists(OUTPUT_ODS):
        print(f'Initial .ods file created: {OUTPUT_ODS}')
    else:
        print(f'ERROR: Expected .ods not found at {OUTPUT_ODS}')

    # GUI-ready startup: open the .ods in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT_ODS}"', delay_sec=3.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
