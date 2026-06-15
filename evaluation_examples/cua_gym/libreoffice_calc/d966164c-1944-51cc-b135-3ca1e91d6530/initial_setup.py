"""
Initial Setup: HR Performance Review Tracker
Task ID: calc_hr_conditional_overdue_reviews_016
Domain: libreoffice_calc

Creates a Reviews sheet with 97 employee records (rows 2-98).
Headers: Emp ID, Name, Department, Review Due Date, Status
No conditional formatting in initial file.
"""

import openpyxl
from datetime import date, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_conditional_overdue_reviews_016'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Reviews'

    # --- Headers ---
    headers = ['Emp ID', 'Name', 'Department', 'Review Due Date', 'Status']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic employee data
    # today = date used when script runs on VM (~2026-03-04)
    # We use fixed dates to ensure reproducibility:
    # past dates (overdue), near-future (within 7 days), far-future
    # Reference: script run date on VM is 2026-03-04

    employees = [
        # (emp_id, name, department, due_date_offset_days, status)
        # Negative = past (overdue if Pending), 0-7 = due soon if Pending
        ('EMP001', 'Sarah Chen', 'Engineering', -45, 'Pending'),
        ('EMP002', 'Marcus Johnson', 'Marketing', -30, 'Completed'),
        ('EMP003', 'Priya Patel', 'Finance', -20, 'Pending'),
        ('EMP004', 'James O\'Brien', 'HR', -15, 'In Progress'),
        ('EMP005', 'Linda Nguyen', 'Engineering', -10, 'Pending'),
        ('EMP006', 'Carlos Mendez', 'Sales', -8, 'Completed'),
        ('EMP007', 'Rachel Kim', 'Marketing', -5, 'Pending'),
        ('EMP008', 'David Okonkwo', 'IT', -3, 'Pending'),
        ('EMP009', 'Fatima Al-Hassan', 'Finance', -1, 'Pending'),
        ('EMP010', 'Brian Sullivan', 'Operations', 0, 'Pending'),
        ('EMP011', 'Yuki Tanaka', 'Engineering', 1, 'Pending'),
        ('EMP012', 'Amara Diallo', 'HR', 2, 'Pending'),
        ('EMP013', 'Connor Walsh', 'Sales', 3, 'Pending'),
        ('EMP014', 'Mei-Ling Zhou', 'IT', 4, 'In Progress'),
        ('EMP015', 'Roberto Garcia', 'Finance', 5, 'Pending'),
        ('EMP016', 'Sandra Kowalski', 'Marketing', 6, 'Pending'),
        ('EMP017', 'Olusegun Adeyemi', 'Engineering', 7, 'Pending'),
        ('EMP018', 'Natasha Ivanova', 'Operations', 8, 'Completed'),
        ('EMP019', 'Ahmed Khalil', 'Sales', 10, 'Pending'),
        ('EMP020', 'Helen Fitzgerald', 'HR', 12, 'Pending'),
        ('EMP021', 'Kwame Asante', 'IT', 14, 'In Progress'),
        ('EMP022', 'Sophia Romano', 'Finance', 15, 'Pending'),
        ('EMP023', 'Tyler Brooks', 'Engineering', 18, 'Pending'),
        ('EMP024', 'Ingrid Larsson', 'Marketing', 20, 'Completed'),
        ('EMP025', 'Patrick Nkosi', 'Sales', 21, 'Pending'),
        ('EMP026', 'Mei Fujimoto', 'Operations', 25, 'Pending'),
        ('EMP027', 'Daniel Reyes', 'HR', 28, 'Pending'),
        ('EMP028', 'Aisha Bello', 'Finance', 30, 'In Progress'),
        ('EMP029', 'Ethan Clarke', 'IT', 32, 'Pending'),
        ('EMP030', 'Vera Petrov', 'Engineering', 35, 'Pending'),
        ('EMP031', 'Jorge Castillo', 'Marketing', -60, 'Pending'),
        ('EMP032', 'Nicole Osei', 'Sales', -55, 'Completed'),
        ('EMP033', 'Michael Chang', 'Finance', -50, 'Pending'),
        ('EMP034', 'Lena Hoffmann', 'HR', -48, 'In Progress'),
        ('EMP035', 'Rashid Omar', 'Engineering', -40, 'Pending'),
        ('EMP036', 'Camille Dupont', 'Operations', -38, 'Completed'),
        ('EMP037', 'Takeshi Mori', 'IT', -35, 'Pending'),
        ('EMP038', 'Blessing Eze', 'Sales', -32, 'Pending'),
        ('EMP039', 'Anna Koroleva', 'Marketing', -28, 'Pending'),
        ('EMP040', 'Fernando Lima', 'Finance', -25, 'In Progress'),
        ('EMP041', 'Chloe Beaumont', 'HR', -22, 'Pending'),
        ('EMP042', 'Samuel Adkins', 'Engineering', -18, 'Completed'),
        ('EMP043', 'Zara Malik', 'IT', -16, 'Pending'),
        ('EMP044', 'Oliver Stanton', 'Sales', -12, 'Pending'),
        ('EMP045', 'Hyun-Ji Park', 'Marketing', -9, 'In Progress'),
        ('EMP046', 'Chioma Obi', 'Operations', -7, 'Pending'),
        ('EMP047', 'Vladimir Sokolov', 'Finance', -4, 'Pending'),
        ('EMP048', 'Isabelle Laurent', 'Engineering', -2, 'Completed'),
        ('EMP049', 'Taiwo Afolabi', 'HR', -1, 'Pending'),
        ('EMP050', 'William Hayes', 'IT', 0, 'In Progress'),
        ('EMP051', 'Amelia Thornton', 'Sales', 1, 'Pending'),
        ('EMP052', 'Bongani Mokoena', 'Marketing', 2, 'Pending'),
        ('EMP053', 'Grace Yamamoto', 'Finance', 3, 'Completed'),
        ('EMP054', 'Hugo Fernandez', 'Engineering', 4, 'Pending'),
        ('EMP055', 'Lily Nakamura', 'HR', 5, 'Pending'),
        ('EMP056', 'Emeka Ike', 'Operations', 6, 'In Progress'),
        ('EMP057', 'Sophie Richter', 'IT', 7, 'Pending'),
        ('EMP058', 'Jomo Kenyatta Jr.', 'Sales', 9, 'Pending'),
        ('EMP059', 'Beatriz Santos', 'Marketing', 11, 'Completed'),
        ('EMP060', 'Lars Eriksson', 'Finance', 13, 'Pending'),
        ('EMP061', 'Nadia Volkov', 'Engineering', 16, 'Pending'),
        ('EMP062', 'Kwesi Mensah', 'HR', 17, 'In Progress'),
        ('EMP063', 'Tanya Oduya', 'IT', 19, 'Pending'),
        ('EMP064', 'Marco Pellegrini', 'Sales', 22, 'Pending'),
        ('EMP065', 'Siobhan Murphy', 'Marketing', 24, 'Completed'),
        ('EMP066', 'Arjun Sharma', 'Operations', 26, 'Pending'),
        ('EMP067', 'Valeria Cruz', 'Finance', 29, 'Pending'),
        ('EMP068', 'Desmond Achebe', 'Engineering', 31, 'In Progress'),
        ('EMP069', 'Elise Moreau', 'HR', 33, 'Pending'),
        ('EMP070', 'Sanjay Menon', 'IT', 36, 'Pending'),
        ('EMP071', 'Ayasha Yellowbird', 'Sales', -90, 'Pending'),
        ('EMP072', 'Bernard Osei', 'Marketing', -85, 'Completed'),
        ('EMP073', 'Carmen Vasquez', 'Finance', -80, 'Pending'),
        ('EMP074', 'Dmitri Lebedev', 'HR', -75, 'In Progress'),
        ('EMP075', 'Ekaterina Morozova', 'Engineering', -70, 'Pending'),
        ('EMP076', 'Fuad Al-Amin', 'Operations', -65, 'Completed'),
        ('EMP077', 'Gabby Thompson', 'IT', -62, 'Pending'),
        ('EMP078', 'Hiroshi Watanabe', 'Sales', -58, 'Pending'),
        ('EMP079', 'Imelda Santos', 'Marketing', -53, 'In Progress'),
        ('EMP080', 'Jabari Freeman', 'Finance', -47, 'Pending'),
        ('EMP081', 'Karin Johansson', 'Engineering', -42, 'Pending'),
        ('EMP082', 'Liam O\'Connor', 'HR', -36, 'Completed'),
        ('EMP083', 'Miriam Goldstein', 'IT', -33, 'Pending'),
        ('EMP084', 'Nnamdi Okafor', 'Sales', -27, 'Pending'),
        ('EMP085', 'Ophelia Marchand', 'Marketing', -24, 'In Progress'),
        ('EMP086', 'Pascal Fontaine', 'Finance', -19, 'Pending'),
        ('EMP087', 'Quincy Washington', 'Engineering', -14, 'Pending'),
        ('EMP088', 'Rhea Krishnamurthy', 'HR', -11, 'Completed'),
        ('EMP089', 'Stefan Müller', 'Operations', -6, 'Pending'),
        ('EMP090', 'Tamara Jackson', 'IT', -2, 'Pending'),
        ('EMP091', 'Ulrich Schneider', 'Sales', 1, 'In Progress'),
        ('EMP092', 'Vanessa Mbeki', 'Marketing', 2, 'Pending'),
        ('EMP093', 'Winston Osei-Bonsu', 'Finance', 3, 'Pending'),
        ('EMP094', 'Xiao Wei', 'Engineering', 5, 'Completed'),
        ('EMP095', 'Yolanda Ferreira', 'HR', 6, 'Pending'),
        ('EMP096', 'Zuberi Makwela', 'Operations', 7, 'Pending'),
        ('EMP097', 'Ama Asare', 'IT', 40, 'Pending'),
    ]

    # Reference date: 2026-03-04 (VM date)
    ref_date = date(2026, 3, 4)

    for row_idx, (emp_id, name, dept, offset, status) in enumerate(employees, 2):
        due_date = ref_date + timedelta(days=offset)
        ws.cell(row=row_idx, column=1, value=emp_id)
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=dept)
        ws.cell(row=row_idx, column=4, value=due_date)
        ws.cell(row=row_idx, column=5, value=status)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Reviews')
    print(f'  Rows: 97 employee records (rows 2-98)')
    print(f'  No conditional formatting applied')

create_initial()
