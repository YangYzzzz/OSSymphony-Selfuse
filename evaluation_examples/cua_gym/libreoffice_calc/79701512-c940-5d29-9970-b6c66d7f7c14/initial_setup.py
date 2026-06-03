"""
Initial Setup: Download PDF from research papers spreadsheet and check citation
Task ID: osworld_multi_apps_pdf_download_cite_003
Domain: libreoffice_calc (multi-app: Chrome, LibreOffice Writer)

Creates:
  - /home/user/research_papers.xlsx on Desktop with paper list including PDF URLs
  - Opens LibreOffice Calc with the spreadsheet
  - Opens Chrome browser (for downloading PDF)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_pdf_download_cite_003'
DESKTOP = '/home/user/Desktop'
OUTPUT = f'{DESKTOP}/research_papers.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    # Ensure Desktop directory exists
    os.makedirs(DESKTOP, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Research Papers"

    # --- Header row ---
    headers = ['Title', 'PDF URL', 'Notes']
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=12, color="FFFFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Data rows ---
    data = [
        [
            'Playing Atari with Deep Reinforcement Learning',
            'https://arxiv.org/pdf/1312.5602',
            'Mnih et al., 2013. DeepMind workshop paper introducing DQN on Atari games.'
        ],
        [
            'Human-level control through deep reinforcement learning',
            'https://www.nature.com/articles/nature14236',
            'Mnih et al., 2015. Nature paper expanding DQN to 49 Atari games.'
        ],
        [
            'Deep Reinforcement Learning with Double Q-learning',
            'https://arxiv.org/pdf/1509.06461',
            'van Hasselt et al., 2016. Addresses overestimation bias in DQN.'
        ],
        [
            'Dueling Network Architectures for Deep Reinforcement Learning',
            'https://arxiv.org/pdf/1511.06581',
            'Wang et al., 2016. Separate value and advantage streams in DQN.'
        ],
        [
            'Prioritized Experience Replay',
            'https://arxiv.org/pdf/1511.05952',
            'Schaul et al., 2016. Sample important transitions more frequently.'
        ],
    ]

    normal_font = Font(size=11)
    url_font = Font(size=11, color="FF0000FF", underline="single")
    wrap_align = Alignment(vertical="top", wrap_text=True)
    url_align = Alignment(vertical="top", wrap_text=False)

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 2:
                cell.font = url_font
                cell.alignment = url_align
            else:
                cell.font = normal_font
                cell.alignment = wrap_align

    # --- Column widths ---
    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 60

    # --- Row heights ---
    ws.row_dimensions[1].height = 22
    for r in range(2, 7):
        ws.row_dimensions[r].height = 40

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Kill any existing LibreOffice processes for clean startup
    subprocess.run(['pkill', '-f', 'soffice'], capture_output=True)
    time.sleep(1.5)

    # GUI-ready startup: open spreadsheet in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.5)

    # Open Chrome browser (needed for downloading PDF)
    launch_gui('google-chrome', delay_sec=2.0)

    print('GUI_READY: launched LibreOffice Calc and Chrome with DISPLAY=:0')


create_initial()
