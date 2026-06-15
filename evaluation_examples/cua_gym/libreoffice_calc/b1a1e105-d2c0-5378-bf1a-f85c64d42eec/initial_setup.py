"""
Initial Setup: Travel expense tracker with exchange rates lookup table
Task ID: calc_gen_personal_029
Domain: libreoffice_calc

Creates:
  - Sheet 'ExchangeRates': currency codes and exchange rates to USD
  - Sheet 'Expenses': headers only + sample raw data rows (no formulas, no validation, no formatting)
"""

import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_personal_029'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Expenses ---
    ws_exp = wb.active
    ws_exp.title = 'Expenses'

    # Headers (row 1)
    headers = ['Date', 'Description', 'Currency', 'Amount', 'USD Amount', 'Category']
    for col, h in enumerate(headers, 1):
        ws_exp.cell(row=1, column=col, value=h)
        ws_exp.cell(row=1, column=col).font = Font(bold=True)

    # Sample expense data — realistic business trip entries
    # NO formulas in USD Amount (E), NO category validation
    expense_data = [
        ['2025-06-02', 'Airport taxi to hotel', 'EUR', 38.50, None, None],
        ['2025-06-02', 'Flight dinner upgrade', 'EUR', 22.00, None, None],
        ['2025-06-03', 'Client lunch at Maison Rouge', 'EUR', 95.40, None, None],
        ['2025-06-03', 'Metro day pass', 'EUR', 7.80, None, None],
        ['2025-06-03', 'Hotel stay night 1', 'EUR', 189.00, None, None],
        ['2025-06-04', 'Conference registration', 'GBP', 320.00, None, None],
        ['2025-06-04', 'Team dinner at The Ivy', 'GBP', 147.60, None, None],
        ['2025-06-04', 'Hotel stay night 2', 'GBP', 215.00, None, None],
        ['2025-06-05', 'Office supplies — printer cartridge', 'SGD', 42.90, None, None],
        ['2025-06-05', 'Client entertainment — museum tour', 'SGD', 88.00, None, None],
        ['2025-06-06', 'Cab to airport', 'SGD', 35.00, None, None],
        ['2025-06-06', 'Tokyo airport lounge', 'JPY', 3500, None, None],
        ['2025-06-07', 'Hotel in Tokyo night 1', 'JPY', 18000, None, None],
        ['2025-06-07', 'Convenience store meals', 'JPY', 1250, None, None],
        ['2025-06-08', 'Train JR Pass top-up', 'JPY', 6500, None, None],
    ]

    for r, row_data in enumerate(expense_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_exp.cell(row=r, column=c, value=val)

    # Adjust column widths for readability
    ws_exp.column_dimensions['A'].width = 13
    ws_exp.column_dimensions['B'].width = 38
    ws_exp.column_dimensions['C'].width = 11
    ws_exp.column_dimensions['D'].width = 12
    ws_exp.column_dimensions['E'].width = 14
    ws_exp.column_dimensions['F'].width = 16

    # --- Sheet 2: ExchangeRates ---
    ws_fx = wb.create_sheet('ExchangeRates')

    # Headers
    ws_fx['A1'] = 'Currency'
    ws_fx['B1'] = 'Rate to USD'
    ws_fx['A1'].font = Font(bold=True)
    ws_fx['B1'].font = Font(bold=True)

    # Exchange rates as specified in context
    fx_data = [
        ('EUR', 1.08),
        ('GBP', 1.27),
        ('JPY', 0.0067),
        ('CAD', 0.74),
        ('AUD', 0.65),
        ('SGD', 0.74),
        ('MXN', 0.058),
    ]

    for r, (currency, rate) in enumerate(fx_data, 2):
        ws_fx.cell(row=r, column=1, value=currency)
        ws_fx.cell(row=r, column=2, value=rate)

    ws_fx.column_dimensions['A'].width = 12
    ws_fx.column_dimensions['B'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheets: {wb.sheetnames}')
    print(f'  Expenses rows (data): {len(expense_data)}')
    print(f'  ExchangeRates rows: {len(fx_data)}')


create_initial()
