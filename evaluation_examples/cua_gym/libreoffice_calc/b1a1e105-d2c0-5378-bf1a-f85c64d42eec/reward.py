"""
Reward Script: Travel expense tracker with currency conversion, dropdowns, and category summary
Task ID: calc_gen_personal_029
Domain: libreoffice_calc
Scoring:
  - Component 1: VLOOKUP formula in E column for USD conversion (0.35 pts)
  - Component 2: Currency data validation dropdown on column C (0.20 pts)
  - Component 3: Category data validation dropdown on column F (0.20 pts)
  - Component 4: Summary section with SUMIF formulas by category (0.15 pts)
  - Component 5: E column number format as $#,##0.00 (0.10 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_gen_personal_029'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Expenses sheet must exist
    if 'Expenses' not in wb.sheetnames:
        print("FAIL: 'Expenses' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Expenses']

    # Component 1: VLOOKUP formula in E column for USD conversion (0.35 pts)
    # Task requires: E2 and down: =D2*VLOOKUP(C2,ExchangeRates.$A:$B,2,0)
    # Initial file has no formulas in column E — this must be added by the agent.
    try:
        vlookup_count = 0
        vlookup_correct = 0
        # Check E2 through E16 (15 data rows in golden file)
        for row in range(2, 17):
            cell_e = ws.cell(row=row, column=5)  # Column E
            cell_d = ws.cell(row=row, column=4)  # Column D (Amount)
            # Only check rows that have data in D column
            if cell_d.value is not None:
                vlookup_count += 1
                val = cell_e.value
                if val and isinstance(val, str):
                    val_upper = val.upper().replace(' ', '')
                    # Check for VLOOKUP referencing ExchangeRates
                    if 'VLOOKUP' in val_upper and 'EXCHANGERATES' in val_upper:
                        vlookup_correct += 1

        if vlookup_count > 0 and vlookup_correct == vlookup_count:
            print(f"PASS: Component 1 — VLOOKUP formula in all {vlookup_correct} data rows of column E (0.35 pts)")
            total_score += 0.35
        elif vlookup_correct > 0:
            partial = round(0.35 * vlookup_correct / vlookup_count, 2)
            print(f"PARTIAL: Component 1 — VLOOKUP formula in {vlookup_correct}/{vlookup_count} data rows ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No VLOOKUP formula found in column E (expected {vlookup_count} rows with VLOOKUP)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Currency data validation dropdown on column C (0.20 pts)
    # Task requires: C column with data validation dropdown for currency codes from ExchangeRates
    # Initial file has NO data validations at all — this must be added by the agent.
    try:
        currency_dv_found = False
        for dv in ws.data_validations.dataValidation:
            if dv.type == 'list':
                sqref_str = str(dv.sqref)
                formula = str(dv.formula1) if dv.formula1 else ''
                # Check if this DV is on column C
                if 'C' in sqref_str:
                    # Check formula references ExchangeRates or contains currency codes
                    if 'ExchangeRates' in formula or 'EXCHANGERATES' in formula.upper():
                        currency_dv_found = True
                        print(f"PASS: Component 2 — Currency dropdown on column C referencing ExchangeRates (formula={formula}) (0.20 pts)")
                        total_score += 0.20
                        break
        if not currency_dv_found:
            print(f"FAIL: Component 2 — No currency dropdown found on column C referencing ExchangeRates")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Category data validation dropdown on column F (0.20 pts)
    # Task requires: F column with dropdown: Meals, Transport, Hotel, Entertainment, Office, Other
    # Initial file has NO data validations — this must be added by the agent.
    try:
        expected_categories = {'Meals', 'Transport', 'Hotel', 'Entertainment', 'Office', 'Other'}
        category_dv_found = False
        for dv in ws.data_validations.dataValidation:
            if dv.type == 'list':
                sqref_str = str(dv.sqref)
                formula = str(dv.formula1) if dv.formula1 else ''
                # Check if this DV is on column F
                if 'F' in sqref_str:
                    # Check that formula contains expected category values
                    categories_in_formula = set()
                    for cat in expected_categories:
                        if cat in formula:
                            categories_in_formula.add(cat)
                    if len(categories_in_formula) >= 5:
                        category_dv_found = True
                        print(f"PASS: Component 3 — Category dropdown on column F with {len(categories_in_formula)} categories (0.20 pts)")
                        total_score += 0.20
                        break
                    elif len(categories_in_formula) > 0:
                        print(f"PARTIAL: Component 3 — Category dropdown on column F with only {len(categories_in_formula)}/6 expected categories")
        if not category_dv_found:
            print(f"FAIL: Component 3 — No category dropdown found on column F with expected categories")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Summary section with SUMIF formulas by category (0.15 pts)
    # Task requires: summary section with SUMIF totaling USD amounts by category
    # Initial file has no summary section — this must be added.
    try:
        sumif_count = 0
        expected_categories_list = ['Meals', 'Transport', 'Hotel', 'Entertainment', 'Office', 'Other']
        # Search for SUMIF formulas in the worksheet
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and 'SUMIF' in cell.value.upper():
                    sumif_count += 1

        # Also check for category labels in summary area
        category_labels_found = 0
        for row in ws.iter_rows(min_row=15, max_row=50):
            for cell in row:
                if cell.value and str(cell.value).strip() in expected_categories_list:
                    category_labels_found += 1

        if sumif_count >= 6 and category_labels_found >= 5:
            print(f"PASS: Component 4 — Summary section with {sumif_count} SUMIF formulas and {category_labels_found} category labels (0.15 pts)")
            total_score += 0.15
        elif sumif_count >= 3:
            partial = 0.08
            print(f"PARTIAL: Component 4 — Summary with {sumif_count} SUMIF formulas, {category_labels_found} category labels ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Insufficient summary section: {sumif_count} SUMIF formulas, {category_labels_found} category labels")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: E column number format as $#,##0.00 (0.10 pts)
    # Task requires: Format E column as $#,##0.00
    # Initial file has no number format on column E data cells.
    try:
        formatted_count = 0
        total_data_cells = 0
        for row in range(2, 17):
            cell_d = ws.cell(row=row, column=4)
            cell_e = ws.cell(row=row, column=5)
            if cell_d.value is not None:
                total_data_cells += 1
                if cell_e.number_format and '$' in cell_e.number_format:
                    formatted_count += 1

        if total_data_cells > 0 and formatted_count == total_data_cells:
            print(f"PASS: Component 5 — All {formatted_count} E column data cells have $#,##0.00 format (0.10 pts)")
            total_score += 0.10
        elif formatted_count > 0:
            partial = round(0.10 * formatted_count / total_data_cells, 2)
            print(f"PARTIAL: Component 5 — {formatted_count}/{total_data_cells} E column cells formatted ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — No currency format found in column E (expected $#,##0.00)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
