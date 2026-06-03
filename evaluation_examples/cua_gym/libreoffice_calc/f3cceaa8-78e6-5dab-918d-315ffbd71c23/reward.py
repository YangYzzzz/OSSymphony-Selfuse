"""
Reward Script: Facilities Maintenance Request Tracker Setup
Task ID: calc_ops_facility_maintenance_requests_033
Domain: libreoffice_calc
Scoring:
  Component 1: Priority dropdown (E2:E61) - Emergency, High, Medium, Low  (0.20 pts)
  Component 2: Status dropdown (G2:G61) - Open, In Progress, On Hold, Closed  (0.20 pts)
  Component 3: Days Open formulas in I2:I61 (0.20 pts)
  Component 4: SLA Target formulas in J2:J61 (0.20 pts)
  Component 5: SLA Status formulas in K2:K61 (0.15 pts)
  Component 6: Red conditional formatting on OVERDUE rows (0.05 pts)
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_facility_maintenance_requests_033'


def normalize_formula(f):
    """Normalize formula for comparison: strip whitespace, upper-case."""
    if not isinstance(f, str):
        return ''
    return f.strip().upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns float between 0.0 and 1.0.
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet 'MaintenanceRequests' must exist
    if 'MaintenanceRequests' not in wb.sheetnames:
        print("CRITICAL: Sheet 'MaintenanceRequests' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['MaintenanceRequests']

    # -----------------------------------------------------------------------
    # Component 1: Data validation dropdown on E2:E61 for Priority
    # Values must be: Emergency, High, Medium, Low
    # (0.20 points)
    # -----------------------------------------------------------------------
    try:
        priority_dv_found = False
        expected_priority_values = {'Emergency', 'High', 'Medium', 'Low'}

        for dv in ws.data_validations.dataValidation:
            if dv.type != 'list':
                continue
            sqref_str = str(dv.sqref)
            # Check if this validation covers E column rows 2-61
            # The sqref may be 'E2:E61' or contain that range
            if 'E' not in sqref_str.upper():
                continue
            # Parse the formula1 to get list items
            formula = dv.formula1 or ''
            # Remove surrounding quotes if present: '"Emergency,High,Medium,Low"'
            formula_clean = formula.strip('"').strip("'")
            items = set(item.strip() for item in formula_clean.split(','))
            # Check all 4 priority values are present
            if expected_priority_values.issubset(items) or items == expected_priority_values:
                # Verify range covers E2:E61
                if 'E2' in sqref_str or 'E2:E61' in sqref_str:
                    priority_dv_found = True
                    print(f"PASS: Component 1 — Priority dropdown on E2:E61 found with values: {formula_clean} (0.20 pts)")
                    total_score += 0.20
                    break
            elif expected_priority_values.issubset(items):
                priority_dv_found = True
                print(f"PASS: Component 1 — Priority dropdown on E2:E61 found with values: {formula_clean} (0.20 pts)")
                total_score += 0.20
                break

        if not priority_dv_found:
            # Report what was found
            found_dvs = [(str(dv.sqref), dv.formula1) for dv in ws.data_validations.dataValidation]
            print(f"FAIL: Component 1 — Priority dropdown not found on E2:E61. Found validations: {found_dvs}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Data validation dropdown on G2:G61 for Status
    # Values must be: Open, In Progress, On Hold, Closed
    # (0.20 points)
    # -----------------------------------------------------------------------
    try:
        status_dv_found = False
        expected_status_values = {'Open', 'In Progress', 'On Hold', 'Closed'}

        for dv in ws.data_validations.dataValidation:
            if dv.type != 'list':
                continue
            sqref_str = str(dv.sqref)
            if 'G' not in sqref_str.upper():
                continue
            formula = dv.formula1 or ''
            formula_clean = formula.strip('"').strip("'")
            items = set(item.strip() for item in formula_clean.split(','))
            if expected_status_values.issubset(items) or items == expected_status_values:
                if 'G2' in sqref_str or 'G2:G61' in sqref_str:
                    status_dv_found = True
                    print(f"PASS: Component 2 — Status dropdown on G2:G61 found with values: {formula_clean} (0.20 pts)")
                    total_score += 0.20
                    break
            elif expected_status_values.issubset(items):
                status_dv_found = True
                print(f"PASS: Component 2 — Status dropdown on G2:G61 found with values: {formula_clean} (0.20 pts)")
                total_score += 0.20
                break

        if not status_dv_found:
            found_dvs = [(str(dv.sqref), dv.formula1) for dv in ws.data_validations.dataValidation]
            print(f"FAIL: Component 2 — Status dropdown not found on G2:G61. Found validations: {found_dvs}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Days Open formulas in I2:I61
    # Formula pattern: =IF(G{row}="Closed",H{row}-B{row},TODAY()-B{row})
    # We check all 60 rows match this pattern (case-insensitive, whitespace-normalized)
    # (0.20 points)
    # -----------------------------------------------------------------------
    try:
        days_open_correct = 0
        days_open_total = 60

        for row in range(2, 62):
            cell_val = ws.cell(row=row, column=9).value  # Column I
            if cell_val is None:
                continue
            val_norm = normalize_formula(str(cell_val))
            # Build expected pattern
            expected = f'=IF(G{row}="CLOSED",H{row}-B{row},TODAY()-B{row})'
            expected_norm = normalize_formula(expected)
            if val_norm == expected_norm:
                days_open_correct += 1

        coverage = days_open_correct / days_open_total
        if coverage >= 0.95:
            print(f"PASS: Component 3 — Days Open formulas in I2:I61: {days_open_correct}/{days_open_total} correct (0.20 pts)")
            total_score += 0.20
        elif coverage >= 0.5:
            partial = round(0.20 * coverage, 2)
            print(f"PARTIAL: Component 3 — Days Open formulas: {days_open_correct}/{days_open_total} rows correct ({partial} pts)")
            total_score += partial
        else:
            # Sample a failing row to report
            sample_val = ws.cell(row=2, column=9).value
            print(f"FAIL: Component 3 — Days Open formulas incomplete: {days_open_correct}/{days_open_total}. I2={repr(sample_val)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: SLA Target formulas in J2:J61
    # Formula pattern: =IF(E{row}="Emergency",1,IF(E{row}="High",3,IF(E{row}="Medium",7,14)))
    # (0.20 points)
    # -----------------------------------------------------------------------
    try:
        sla_target_correct = 0
        sla_target_total = 60

        for row in range(2, 62):
            cell_val = ws.cell(row=row, column=10).value  # Column J
            if cell_val is None:
                continue
            val_norm = normalize_formula(str(cell_val))
            expected = f'=IF(E{row}="EMERGENCY",1,IF(E{row}="HIGH",3,IF(E{row}="MEDIUM",7,14)))'
            expected_norm = normalize_formula(expected)
            if val_norm == expected_norm:
                sla_target_correct += 1

        coverage = sla_target_correct / sla_target_total
        if coverage >= 0.95:
            print(f"PASS: Component 4 — SLA Target formulas in J2:J61: {sla_target_correct}/{sla_target_total} correct (0.20 pts)")
            total_score += 0.20
        elif coverage >= 0.5:
            partial = round(0.20 * coverage, 2)
            print(f"PARTIAL: Component 4 — SLA Target formulas: {sla_target_correct}/{sla_target_total} rows correct ({partial} pts)")
            total_score += partial
        else:
            sample_val = ws.cell(row=2, column=10).value
            print(f"FAIL: Component 4 — SLA Target formulas incomplete: {sla_target_correct}/{sla_target_total}. J2={repr(sample_val)}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -----------------------------------------------------------------------
    # Component 5: SLA Status formulas in K2:K61
    # Formula pattern: =IF(AND(I{row}>J{row},G{row}<>"Closed"),"OVERDUE",IF(AND(G{row}="Closed",I{row}<=J{row}),"MET","OPEN"))
    # (0.15 points)
    # -----------------------------------------------------------------------
    try:
        sla_status_correct = 0
        sla_status_total = 60

        for row in range(2, 62):
            cell_val = ws.cell(row=row, column=11).value  # Column K
            if cell_val is None:
                continue
            val_norm = normalize_formula(str(cell_val))
            expected = (
                f'=IF(AND(I{row}>J{row},G{row}<>"CLOSED"),"OVERDUE",'
                f'IF(AND(G{row}="CLOSED",I{row}<=J{row}),"MET","OPEN"))'
            )
            expected_norm = normalize_formula(expected)
            if val_norm == expected_norm:
                sla_status_correct += 1

        coverage = sla_status_correct / sla_status_total
        if coverage >= 0.95:
            print(f"PASS: Component 5 — SLA Status formulas in K2:K61: {sla_status_correct}/{sla_status_total} correct (0.15 pts)")
            total_score += 0.15
        elif coverage >= 0.5:
            partial = round(0.15 * coverage, 2)
            print(f"PARTIAL: Component 5 — SLA Status formulas: {sla_status_correct}/{sla_status_total} rows correct ({partial} pts)")
            total_score += partial
        else:
            sample_val = ws.cell(row=2, column=11).value
            print(f"FAIL: Component 5 — SLA Status formulas incomplete: {sla_status_correct}/{sla_status_total}. K2={repr(sample_val)}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # -----------------------------------------------------------------------
    # Component 6: Conditional formatting — red fill for OVERDUE rows
    # Rule: formula '$K2="OVERDUE"' applies red fill to A2:K61 range
    # (0.05 points)
    # -----------------------------------------------------------------------
    try:
        cf_found = False
        for cf in ws.conditional_formatting:
            for rule in cf.rules:
                if rule.type not in ('expression', 'formula'):
                    continue
                formulas = rule.formula if hasattr(rule, 'formula') and rule.formula else []
                for formula in formulas:
                    formula_upper = formula.upper().replace(' ', '')
                    # Look for OVERDUE pattern: $K*="OVERDUE" or K*="OVERDUE"
                    if 'OVERDUE' in formula_upper and 'K' in formula_upper:
                        # Check for red fill
                        if hasattr(rule, 'dxf') and rule.dxf and hasattr(rule.dxf, 'fill') and rule.dxf.fill:
                            try:
                                fill_color = rule.dxf.fill.fgColor.rgb
                                # Accept any shade of red: FF0000, FFFF0000
                                if fill_color and ('FF0000' in fill_color.upper() or fill_color.upper() in ('FFFF0000',)):
                                    cf_found = True
                                    print(f"PASS: Component 6 — Red conditional formatting for OVERDUE rows found (fill={fill_color}) (0.05 pts)")
                                    total_score += 0.05
                                    break
                            except Exception:
                                pass
                if cf_found:
                    break
            if cf_found:
                break

        if not cf_found:
            print("FAIL: Component 6 — Red conditional formatting for OVERDUE rows not found")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
