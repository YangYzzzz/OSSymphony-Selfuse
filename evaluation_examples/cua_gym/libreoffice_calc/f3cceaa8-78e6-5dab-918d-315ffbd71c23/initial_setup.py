"""
Initial Setup: Facilities Maintenance Request Tracker
Task ID: calc_ops_facility_maintenance_requests_033
Domain: libreoffice_calc

Creates a spreadsheet with 60 maintenance requests. Columns I (Days Open),
J (SLA Target Days), and K (SLA Status) are intentionally left empty.
No data validation dropdowns are present yet.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date, timedelta
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_facility_maintenance_requests_033'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'MaintenanceRequests'

    # --- Headers ---
    headers = [
        'Request ID', 'Date Raised', 'Location', 'Description',
        'Priority', 'Assigned To', 'Status', 'Date Closed',
        'Days Open', 'SLA Target Days', 'SLA Status'
    ]
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Column widths ---
    col_widths = [14, 14, 22, 48, 12, 18, 14, 14, 12, 16, 12]
    col_letters = ['A','B','C','D','E','F','G','H','I','J','K']
    for letter, width in zip(col_letters, col_widths):
        ws.column_dimensions[letter].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    # --- Realistic Data ---
    locations = [
        'Building A - Floor 1', 'Building A - Floor 2', 'Building A - Floor 3',
        'Building B - Floor 1', 'Building B - Floor 2', 'Building C - Lobby',
        'Building C - Basement', 'Parking Lot North', 'Parking Lot South',
        'Server Room 1', 'Server Room 2', 'Cafeteria', 'Conference Room 101',
        'Conference Room 202', 'Reception Area', 'Loading Dock', 'Rooftop',
        'Stairwell A', 'Stairwell B', 'Gym/Fitness Center'
    ]

    descriptions = [
        'HVAC unit not cooling properly in the east wing',
        'Leaking pipe under sink in restroom',
        'Broken window latch on third floor office',
        'Elevator door sensor malfunctioning',
        'Flickering fluorescent lights in hallway',
        'Fire extinguisher needs recharging',
        'Broken door handle on stairwell door',
        'Water stain on ceiling tiles from roof leak',
        'Damaged electrical outlet sparking',
        'Generator backup not starting during test',
        'Broken security camera near parking entrance',
        'Sewage odor in basement corridor',
        'Cracked floor tiles near lobby entrance',
        'Air vent blocked causing temperature issues',
        'Pest infestation reported in storage room',
        'Exterior lighting out along north pathway',
        'Loading dock bay door not closing fully',
        'Emergency exit sign not illuminated',
        'Water heater making unusual noise',
        'AC thermostat unresponsive in conference room',
        'Graffiti on south exterior wall',
        'Roof drain clogged causing pooling water',
        'Broken blind in executive office 4B',
        'Intercom system not working on 2nd floor',
        'Carpet torn near stairwell entrance',
        'Ceiling fan wobbling dangerously in lobby',
        'Plumbing blockage in men\'s restroom',
        'Electrical panel breaker keeps tripping',
        'Handrail loose on main staircase',
        'Kitchen refrigerator temperature alarm',
        'Parking lot pothole near entrance gate',
        'Server room cooling fan failure',
        'Broken projector mount in meeting room',
        'Slippery floor near building entrance (no mat)',
        'Gas leak smell near utility room',
        'Damaged insulation in storage area',
        'Garage door opener malfunctioning',
        'Water fountain not working on floor 3',
        'Boiler pressure gauge reading abnormal',
        'Elevator inspection certificate expired',
        'Mold detected in basement restroom',
        'Damaged fire door self-closing mechanism',
        'Exterior gutter overflowing and damaged',
        'Main entrance automatic door stuck open',
        'Backup lighting battery replacement needed',
        'Sprinkler head damaged in storage room',
        'Drainage issue causing flooding in parking',
        'Broken lock on server room door',
        'Noise complaint from HVAC unit roof',
        'Sewage pump alarm triggered in basement',
        'Ceiling tile collapsed in hallway',
        'Power surge damaged equipment in lab',
        'Wheelchair ramp damaged, accessibility issue',
        'Roof flashing deteriorating near vent',
        'Hot water not available on floor 2',
        'Gym treadmill safety cord missing',
        'Loading bay scale giving incorrect readings',
        'External signage light bulb out',
        'Rodent trap placement needed in cafeteria',
        'Security door badge reader not responding',
    ]

    assigned_staff = [
        'Tom Hargreaves', 'Lisa Mendoza', 'Derek Okafor', 'Sandra Patel',
        'Mike Whitfield', 'Rachel Torres', 'Kevin Nguyen', 'Angela Brooks',
        'James Fitzgerald', 'Nina Coleman', 'Paul Stein', 'Carla Diaz'
    ]

    priorities = ['Emergency', 'High', 'Medium', 'Low']
    priority_weights = [0.05, 0.20, 0.45, 0.30]

    statuses_for_closed = ['Closed']
    statuses_for_open = ['Open', 'In Progress', 'On Hold']

    base_date = date(2025, 1, 1)
    today = date(2026, 3, 4)

    random.seed(42)

    for i in range(60):
        row = i + 2
        req_id = f'MR-2025-{str(i+1).zfill(3)}'

        # Date raised: spread over roughly the past 14 months
        days_ago = random.randint(5, 425)
        date_raised = today - timedelta(days=days_ago)
        # Clamp to after base date
        if date_raised < base_date:
            date_raised = base_date + timedelta(days=random.randint(0, 30))

        location = locations[i % len(locations)]
        description = descriptions[i % len(descriptions)]

        priority = random.choices(priorities, weights=priority_weights, k=1)[0]
        assigned_to = random.choice(assigned_staff)

        # Determine status
        if random.random() < 0.35:
            status = 'Closed'
            # Closed date: 1 to 30 days after raised
            close_offset = random.randint(1, 30)
            date_closed = date_raised + timedelta(days=close_offset)
            if date_closed > today:
                date_closed = today
        elif random.random() < 0.3:
            status = 'On Hold'
            date_closed = None
        elif random.random() < 0.5:
            status = 'In Progress'
            date_closed = None
        else:
            status = 'Open'
            date_closed = None

        ws.cell(row=row, column=1, value=req_id)
        cell_b = ws.cell(row=row, column=2, value=date_raised)
        cell_b.number_format = 'yyyy-mm-dd'
        ws.cell(row=row, column=3, value=location)
        ws.cell(row=row, column=4, value=description)
        ws.cell(row=row, column=5, value=priority)
        ws.cell(row=row, column=6, value=assigned_to)
        ws.cell(row=row, column=7, value=status)
        if date_closed is not None:
            cell_h = ws.cell(row=row, column=8, value=date_closed)
            cell_h.number_format = 'yyyy-mm-dd'
        # Columns I (9), J (10), K (11) — intentionally LEFT EMPTY

    # Auto-filter on header row
    ws.auto_filter.ref = 'A1:K1'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: MaintenanceRequests')
    print(f'  Rows: 61 (1 header + 60 data rows)')
    print(f'  Columns: A-K (I, J, K empty — to be filled by agent)')

create_initial()
