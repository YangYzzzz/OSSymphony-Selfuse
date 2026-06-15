"""
Initial Setup: Skills Gap Analysis Matrix
Task ID: calc_hr_077
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_077'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Skills (current competency levels) ---
    ws_skills = wb.active
    ws_skills.title = 'Skills'

    headers = ['Employee', 'Role', 'Excel', 'Communication', 'Leadership', 'Technical', 'Project Mgmt']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    white_font = Font(bold=True, size=11, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws_skills.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Employee data with current competency levels (1-5 scale)
    employees = [
        ['Sarah Chen', 'Analyst', 4, 3, 2, 4, 2],
        ['Marcus Johnson', 'Manager', 3, 4, 4, 3, 5],
        ['Priya Patel', 'Analyst', 5, 3, 1, 5, 2],
        ['James Wilson', 'Director', 3, 5, 5, 2, 4],
        ['Elena Rodriguez', 'Manager', 4, 4, 3, 3, 4],
        ['David Kim', 'Developer', 3, 2, 1, 5, 2],
        ['Amanda Foster', 'Director', 2, 5, 4, 3, 5],
        ['Robert Chang', 'Developer', 4, 2, 2, 5, 1],
        ['Lisa Thompson', 'Analyst', 3, 4, 2, 4, 3],
        ['Michael Brown', 'Manager', 4, 3, 4, 3, 4],
        ['Jennifer Lee', 'Developer', 5, 3, 1, 4, 2],
        ['Carlos Mendez', 'Director', 3, 4, 5, 3, 5],
    ]

    for r, row_data in enumerate(employees, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_skills.cell(row=r, column=c, value=val)
            if c >= 3:  # numeric skill columns
                cell.alignment = Alignment(horizontal="center")

    # Set column widths
    ws_skills.column_dimensions['A'].width = 20
    ws_skills.column_dimensions['B'].width = 14
    for col_letter in ['C', 'D', 'E', 'F', 'G']:
        ws_skills.column_dimensions[col_letter].width = 16

    # --- Sheet 2: Required (required competency levels by role) ---
    ws_req = wb.create_sheet('Required')

    req_headers = ['Role', 'Excel', 'Communication', 'Leadership', 'Technical', 'Project Mgmt']
    for col, h in enumerate(req_headers, 1):
        cell = ws_req.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    required_levels = [
        ['Analyst', 5, 3, 2, 5, 3],
        ['Manager', 4, 5, 5, 3, 5],
        ['Director', 3, 5, 5, 4, 5],
        ['Developer', 5, 3, 2, 5, 3],
    ]

    for r, row_data in enumerate(required_levels, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_req.cell(row=r, column=c, value=val)
            if c >= 2:
                cell.alignment = Alignment(horizontal="center")

    ws_req.column_dimensions['A'].width = 14
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws_req.column_dimensions[col_letter].width = 16

    # --- Sheet 3: GapAnalysis (empty - to be filled by agent) ---
    ws_gap = wb.create_sheet('GapAnalysis')

    gap_headers = ['Employee', 'Role', 'Excel Gap', 'Communication Gap',
                   'Leadership Gap', 'Technical Gap', 'Project Mgmt Gap', 'Priority Score']
    for col, h in enumerate(gap_headers, 1):
        cell = ws_gap.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
        cell.alignment = header_align

    ws_gap.column_dimensions['A'].width = 20
    ws_gap.column_dimensions['B'].width = 14
    for col_letter in ['C', 'D', 'E', 'F', 'G']:
        ws_gap.column_dimensions[col_letter].width = 20
    ws_gap.column_dimensions['H'].width = 16

    # GapAnalysis sheet has ONLY headers - no formulas, no data, no formatting beyond headers

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
