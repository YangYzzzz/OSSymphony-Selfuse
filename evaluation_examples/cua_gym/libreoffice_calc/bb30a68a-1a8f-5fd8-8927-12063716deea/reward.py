"""
Reward Script: Skills Gap Analysis Matrix
Task ID: calc_hr_077
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Employee/Role references in GapAnalysis A2:B13
  Component 2 (0.35): VLOOKUP gap formulas in GapAnalysis C2:G13
  Component 3 (0.20): Priority score formulas in GapAnalysis H2:H13
  Component 4 (0.20): Conditional formatting on C2:G13 for gaps > 2
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_077'


def persist_app_state(domain):
    """Best-effort save of any open LibreOffice document."""
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print(f"PERSIST: ctrl+s sent for {domain}")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: GapAnalysis sheet must exist
    if 'GapAnalysis' not in wb.sheetnames:
        print("CRITICAL: 'GapAnalysis' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['GapAnalysis']

    # -------------------------------------------------------------------------
    # Component 1: Employee/Role references in GapAnalysis A2:B13 (0.25 points)
    # The golden file has formulas like =Skills!A2 and =Skills!B2 linking
    # employee names and roles from the Skills sheet.
    # We accept either a formula referencing Skills sheet OR the resolved value
    # matching the Skills sheet data.
    # -------------------------------------------------------------------------
    try:
        ref_count = 0
        total_refs = 24  # 12 employees * 2 columns (A, B)

        # Also load Skills sheet to cross-check resolved values
        skills_ws = wb['Skills'] if 'Skills' in wb.sheetnames else None

        for row in range(2, 14):
            for col_letter, col_idx in [('A', 1), ('B', 2)]:
                cell = ws.cell(row=row, column=col_idx)
                val = cell.value
                if val is None:
                    continue

                val_str = str(val).strip()

                # Check if it's a formula referencing Skills sheet
                if val_str.startswith('=') and 'SKILLS!' in val_str.upper():
                    ref_count += 1
                elif skills_ws is not None:
                    # Accept resolved value that matches Skills sheet
                    skills_val = skills_ws.cell(row=row, column=col_idx).value
                    if skills_val is not None and str(skills_val).strip() == val_str:
                        ref_count += 1

        ratio = ref_count / total_refs if total_refs > 0 else 0
        if ratio >= 0.8:
            print(f"PASS: Component 1 — Employee/Role references: {ref_count}/{total_refs} valid ({0.25} pts)")
            total_score += 0.25
        elif ratio >= 0.5:
            partial = round(0.25 * ratio, 2)
            print(f"PARTIAL: Component 1 — Employee/Role references: {ref_count}/{total_refs} valid ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Employee/Role references: {ref_count}/{total_refs} valid")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: VLOOKUP gap formulas in C2:G13 (0.35 points)
    # Each cell should contain a VLOOKUP against the Required sheet minus the
    # current skill level. Pattern: =VLOOKUP(...Required...)-Skills!<col><row>
    # We also accept computed numeric gap values that match expected gaps.
    # -------------------------------------------------------------------------
    try:
        vlookup_count = 0
        total_gap_cells = 60  # 12 rows * 5 skill columns

        # Build expected gap values from Skills and Required data
        skills_ws = wb['Skills'] if 'Skills' in wb.sheetnames else None
        req_ws = wb['Required'] if 'Required' in wb.sheetnames else None

        # Build required levels lookup: role -> [Excel, Comm, Lead, Tech, ProjMgmt]
        required_levels = {}
        if req_ws:
            for r in range(2, req_ws.max_row + 1):
                role = req_ws.cell(row=r, column=1).value
                if role:
                    required_levels[str(role).strip()] = [
                        req_ws.cell(row=r, column=c).value for c in range(2, 7)
                    ]

        for row in range(2, 14):
            for col_idx in range(3, 8):  # C=3 through G=7
                cell = ws.cell(row=row, column=col_idx)
                val = cell.value
                if val is None:
                    continue

                val_str = str(val).strip()

                # Check for VLOOKUP formula pattern
                if val_str.startswith('=') and 'VLOOKUP' in val_str.upper():
                    vlookup_count += 1
                elif skills_ws and req_ws and required_levels:
                    # Accept computed gap value
                    role = skills_ws.cell(row=row, column=2).value
                    if role and str(role).strip() in required_levels:
                        skill_idx = col_idx - 3  # 0-based index into required levels
                        req_val = required_levels[str(role).strip()][skill_idx]
                        cur_val = skills_ws.cell(row=row, column=col_idx).value
                        if req_val is not None and cur_val is not None:
                            expected_gap = req_val - cur_val
                            try:
                                if abs(float(val) - expected_gap) < 0.01:
                                    vlookup_count += 1
                            except (ValueError, TypeError):
                                pass

        ratio = vlookup_count / total_gap_cells if total_gap_cells > 0 else 0
        if ratio >= 0.8:
            print(f"PASS: Component 2 — VLOOKUP gap formulas: {vlookup_count}/{total_gap_cells} valid ({0.35} pts)")
            total_score += 0.35
        elif ratio >= 0.5:
            partial = round(0.35 * ratio, 2)
            print(f"PARTIAL: Component 2 — VLOOKUP gap formulas: {vlookup_count}/{total_gap_cells} valid ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — VLOOKUP gap formulas: {vlookup_count}/{total_gap_cells} valid")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Priority score formulas in H2:H13 (0.20 points)
    # Each cell should contain a weighted sum of gaps. The golden uses:
    #   =C*1 + D*1.2 + E*1.5 + F*1.3 + G*1
    # We accept any formula that references gap columns or a computed value
    # that matches the expected weighted sum.
    # -------------------------------------------------------------------------
    try:
        priority_count = 0
        total_priority_cells = 12  # rows 2-13

        # Weights used in golden: Excel*1, Comm*1.2, Lead*1.5, Tech*1.3, ProjMgmt*1
        weights = [1.0, 1.2, 1.5, 1.3, 1.0]

        for row in range(2, 14):
            cell = ws.cell(row=row, column=8)  # column H
            val = cell.value
            if val is None:
                continue

            val_str = str(val).strip()

            # Check for a formula that references gap columns (any weighting scheme)
            if val_str.startswith('=') and any(ref in val_str.upper() for ref in ['C', 'D', 'E', 'F', 'G']):
                priority_count += 1
            else:
                # Accept computed priority value matching the weighted sum
                # Compute expected from Skills + Required data
                if skills_ws and req_ws and required_levels:
                    role = skills_ws.cell(row=row, column=2).value
                    if role and str(role).strip() in required_levels:
                        gaps = []
                        for ci in range(5):
                            req_val = required_levels[str(role).strip()][ci]
                            cur_val = skills_ws.cell(row=row, column=ci + 3).value
                            if req_val is not None and cur_val is not None:
                                gaps.append(req_val - cur_val)
                            else:
                                gaps.append(0)
                        expected_priority = sum(g * w for g, w in zip(gaps, weights))
                        try:
                            if abs(float(val) - expected_priority) < 0.1:
                                priority_count += 1
                        except (ValueError, TypeError):
                            pass

        ratio = priority_count / total_priority_cells if total_priority_cells > 0 else 0
        if ratio >= 0.8:
            print(f"PASS: Component 3 — Priority score formulas: {priority_count}/{total_priority_cells} valid ({0.20} pts)")
            total_score += 0.20
        elif ratio >= 0.5:
            partial = round(0.20 * ratio, 2)
            print(f"PARTIAL: Component 3 — Priority score formulas: {priority_count}/{total_priority_cells} valid ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Priority score formulas: {priority_count}/{total_priority_cells} valid")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: Conditional formatting for gaps > 2 (0.20 points)
    # The task requires cells with gaps > 2 to be highlighted.
    # Golden has: CellIsRule on C2:G13, operator=greaterThan, formula=['2'],
    # red fill (FFFF0000).
    # -------------------------------------------------------------------------
    try:
        cf_rules = ws.conditional_formatting

        def _check_cf_gap_highlight(cf_rules):
            """Return whether any CF rule highlights gaps > 2 on gap columns."""
            for cf in cf_rules:
                for rule in cf.rules:
                    # Check for a rule that triggers on values > 2
                    matched = (
                        (rule.type == 'cellIs' and rule.operator == 'greaterThan'
                         and rule.formula and '2' in str(rule.formula[0]).strip())
                        or
                        (rule.type == 'expression' and rule.formula
                         and any(pat in str(rule.formula[0]).upper() for pat in ['>2', '> 2']))
                    )
                    if matched:
                        # Verify the range covers gap columns (C through G)
                        range_str = str(cf).upper()
                        if any(col in range_str for col in ['C', 'D', 'E', 'F', 'G']):
                            return 1
            return 0

        if _check_cf_gap_highlight(cf_rules):
            print(f"PASS: Component 4 — Conditional formatting for gaps > 2 ({0.20} pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 4 — No conditional formatting found for gaps > 2")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state('libreoffice_calc')

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
