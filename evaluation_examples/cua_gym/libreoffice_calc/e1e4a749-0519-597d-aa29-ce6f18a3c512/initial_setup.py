"""
Initial Setup: TV Series 2022 Spreadsheet with Empty Season 1 Episodes Column
Task ID: osworld_multi_apps_book_reading_rate_006
Domain: libreoffice_calc (multi-app: also uses LibreOffice Writer)

Creates:
  - /home/user/series_2022.xlsx: Spreadsheet with TV series data, Season 1 Episodes empty
  - /home/user/Desktop/most_episodes.docx: Empty document for the agent to write the answer
"""

import os
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from docx import Document

WORKDIR = '/home/user'
DESKTOP = '/home/user/Desktop'
TASK_ID = 'osworld_multi_apps_book_reading_rate_006'
XLSX_OUTPUT = f'{WORKDIR}/series_2022.xlsx'
DOCX_OUTPUT = f'{DESKTOP}/most_episodes.docx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    # --- Create series_2022.xlsx ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'TV Series 2022'

    # Column headers with formatting
    headers = ['Series Title', 'Network', 'Season 1 Episodes']
    header_font = Font(name='Calibri', bold=True, size=12)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # TV series data (realistic 2022 series)
    # Season 1 Episodes column intentionally left EMPTY — agent must look up on IMDB
    series_data = [
        ['Succession', 'HBO', None],
        ['The White Lotus', 'HBO', None],
        ['Euphoria', 'HBO', None],
        ['Yellowstone', 'Paramount Network', None],
        ['House of the Dragon', 'HBO', None],
    ]

    data_font = Font(name='Calibri', size=11)
    for r, row_data in enumerate(series_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font

    # Set column widths for readability
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(XLSX_OUTPUT)
    print(f'Initial file created: {XLSX_OUTPUT}')

    # --- Create empty most_episodes.docx on Desktop ---
    os.makedirs(DESKTOP, exist_ok=True)
    doc = Document()
    # Empty document — agent will write the series name here
    doc.save(DOCX_OUTPUT)
    print(f'Empty document created: {DOCX_OUTPUT}')

    # --- GUI-ready startup: open both files ---
    # Open the spreadsheet in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{XLSX_OUTPUT}"', delay_sec=2.0)
    # Open the empty document in LibreOffice Writer
    launch_gui(f'libreoffice --writer "{DOCX_OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc and Writer with DISPLAY=:0')


create_initial()
