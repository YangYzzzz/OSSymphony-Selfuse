"""
Initial Setup: Set page order for Dashboard sheet
Task ID: calc_mcp_077
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_077'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Dashboard sheet: wide and tall data area ---
    ws = wb.active
    ws.title = 'Dashboard'

    # Headers spanning many columns (wide layout to ensure multiple print pages)
    headers = [
        'Region', 'Product Line', 'Q1 Revenue', 'Q1 Units', 'Q1 Margin',
        'Q2 Revenue', 'Q2 Units', 'Q2 Margin', 'Q3 Revenue', 'Q3 Units',
        'Q3 Margin', 'Q4 Revenue', 'Q4 Units', 'Q4 Margin', 'Annual Total',
        'YoY Growth', 'Market Share', 'Customer Count', 'Avg Order Value', 'Notes'
    ]
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.font = Font(bold=True, size=11, color="FFFFFF")

    # Data rows (tall layout - 30 rows of data)
    regions = ['North America', 'Europe', 'Asia Pacific', 'Latin America', 'Middle East']
    products = ['Enterprise Suite', 'Cloud Platform', 'Data Analytics', 'Security Pro', 'DevOps Tools', 'AI Accelerator']

    import random
    random.seed(42)

    row_idx = 2
    for region in regions:
        for product in products:
            q1_rev = round(random.uniform(150000, 900000), 2)
            q1_units = random.randint(50, 500)
            q1_margin = round(random.uniform(0.15, 0.55), 4)
            q2_rev = round(q1_rev * random.uniform(0.9, 1.15), 2)
            q2_units = random.randint(50, 550)
            q2_margin = round(random.uniform(0.15, 0.55), 4)
            q3_rev = round(q1_rev * random.uniform(0.85, 1.2), 2)
            q3_units = random.randint(45, 520)
            q3_margin = round(random.uniform(0.15, 0.55), 4)
            q4_rev = round(q1_rev * random.uniform(0.95, 1.25), 2)
            q4_units = random.randint(55, 560)
            q4_margin = round(random.uniform(0.15, 0.55), 4)
            annual = round(q1_rev + q2_rev + q3_rev + q4_rev, 2)
            yoy = round(random.uniform(-0.05, 0.25), 4)
            market_share = round(random.uniform(0.02, 0.18), 4)
            cust_count = random.randint(100, 5000)
            avg_order = round(annual / max(q1_units + q2_units + q3_units + q4_units, 1), 2)
            notes = f'Reviewed {random.choice(["Jan", "Feb", "Mar"])} 2025'

            row_data = [
                region, product, q1_rev, q1_units, q1_margin,
                q2_rev, q2_units, q2_margin, q3_rev, q3_units,
                q3_margin, q4_rev, q4_units, q4_margin, annual,
                yoy, market_share, cust_count, avg_order, notes
            ]
            for c, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=c, value=val)
                # Format percentages
                if c in (5, 8, 11, 14, 16, 17):
                    cell.number_format = '0.00%'
                # Format currency
                elif c in (3, 6, 9, 12, 15, 19):
                    cell.number_format = '$#,##0.00'
                # Format integers
                elif c in (4, 7, 10, 13, 18):
                    cell.number_format = '#,##0'
            row_idx += 1

    # Set column widths for readability
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 20
    for col_letter in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']:
        ws.column_dimensions[col_letter].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Page order is default (overThenDown = left to right, then down)
    # Do NOT set ws.page_setup.pageOrder - leave as default

    # --- Summary sheet (secondary) ---
    ws2 = wb.create_sheet('Summary')
    ws2['A1'] = 'Regional Performance Summary'
    ws2['A1'].font = Font(bold=True, size=14)
    ws2['A3'] = 'Region'
    ws2['B3'] = 'Total Revenue'
    ws2['C3'] = 'Avg Margin'
    for i, region in enumerate(regions, 4):
        ws2.cell(row=i, column=1, value=region)
        ws2.cell(row=i, column=2, value=round(random.uniform(2000000, 8000000), 2))
        ws2.cell(row=i, column=2).number_format = '$#,##0.00'
        ws2.cell(row=i, column=3, value=round(random.uniform(0.25, 0.45), 4))
        ws2.cell(row=i, column=3).number_format = '0.00%'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
