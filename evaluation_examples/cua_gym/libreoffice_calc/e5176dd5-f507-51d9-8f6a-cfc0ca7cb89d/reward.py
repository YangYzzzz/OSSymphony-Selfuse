"""
Reward Script: Set page order for 'Dashboard' sheet to 'Top to bottom, then right'
Task ID: calc_mcp_077
Domain: libreoffice_calc
Scoring:
  Component 1 (0.6): pageOrder on Dashboard sheet is 'downThenOver'
  Component 2 (0.4): pageOrder is correct AND sheet data integrity preserved
"""

import os
import time


WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_077'


def persist_app_state(domain: str):
    """
    Save any unsaved GUI edits before verification.
    For page-order tasks, we kill LibreOffice instead of ctrl+s to avoid
    saving unintended intermediate states. The agent should have saved already.
    """
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Dashboard sheet must exist
    if 'Dashboard' not in wb.sheetnames:
        print(f"FAIL: 'Dashboard' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Dashboard']

    # Component 1: pageOrder on Dashboard is 'downThenOver' (0.6 points)
    # Initial state: pageOrder is None (default = overThenDown)
    # Golden state: pageOrder is 'downThenOver'
    try:
        page_order = ws.page_setup.pageOrder
        if page_order == 'downThenOver':
            print(f"PASS: Component 1 — pageOrder is 'downThenOver' (0.6 pts)")
            total_score += 0.6
        else:
            print(f"FAIL: Component 1 — expected pageOrder='downThenOver', found: {page_order}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: pageOrder is correct AND data integrity preserved (0.4 points)
    # This compound check ensures the agent changed the setting without corrupting data.
    # The data integrity part (headers, non-empty rows) is a sub-condition anchored
    # to the pageOrder change — it only awards points if pageOrder is also correct.
    try:
        page_order = ws.page_setup.pageOrder
        if page_order != 'downThenOver':
            print(f"FAIL: Component 2 — pageOrder not set, skipping data integrity check")
        else:
            # Check that expected headers are still present
            expected_headers = ['Region', 'Product Line']
            actual_headers = [ws.cell(row=1, column=c).value for c in range(1, 3)]
            headers_ok = all(
                exp == act for exp, act in zip(expected_headers, actual_headers)
            )
            # Check that data rows exist (at least 10 non-empty rows)
            non_empty_rows = 0
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=1).value is not None:
                    non_empty_rows += 1
            data_ok = non_empty_rows >= 10

            if headers_ok and data_ok:
                print(f"PASS: Component 2 — pageOrder correct + data intact "
                      f"(headers={actual_headers}, data_rows={non_empty_rows}) (0.4 pts)")
                total_score += 0.4
            else:
                print(f"FAIL: Component 2 — data integrity issue: "
                      f"headers_ok={headers_ok} (got {actual_headers}), "
                      f"data_ok={data_ok} (rows={non_empty_rows})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
