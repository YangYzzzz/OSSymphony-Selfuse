"""
Initial Setup: Albums listening history spreadsheet with Chrome open to Rolling Stone
Task ID: osworld_multi_apps_misc_023
Domain: libreoffice_calc (multi-app: Chrome + LibreOffice Calc)

Creates my_albums.xlsx with the user's listening history (columns: Rank, Album, Artist, Year).
Also opens Chrome with Rolling Stone's 500 Greatest Albums list and LibreOffice Calc with the file.
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_misc_023'
OUTPUT = f'{WORKDIR}/my_albums.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: My Albums ---
    ws = wb.active
    ws.title = 'My Albums'

    # Headers: Rank (RS rank), Album, Artist, Year
    headers = ['Rank', 'Album', 'Artist', 'Year']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # User's listening history — a mix of Rolling Stone top-30 albums and other classics.
    # Includes some from RS top 30 (so those won't appear in must_listen)
    # and some albums not in RS top 30.
    # RS top-30 albums already listened to (by rank): 5, 7, 10, 15, 18, 21, 24, 30
    user_albums = [
        # RS top-30 albums the user has already heard
        (5,  'Abbey Road',                          'The Beatles',       1969),
        (7,  'Rumours',                             'Fleetwood Mac',     1977),
        (10, 'The Miseducation of Lauryn Hill',     'Lauryn Hill',       1998),
        (15, 'Thriller',                            'Michael Jackson',   1982),
        (18, 'Born to Run',                         'Bruce Springsteen', 1975),
        (21, 'Led Zeppelin IV',                     'Led Zeppelin',      1971),
        (24, 'Highway 61 Revisited',                'Bob Dylan',         1965),
        (30, 'Kind of Blue',                        'Miles Davis',       1959),
        # Other albums the user listens to (not in RS top 30)
        (None, 'OK Computer',                       'Radiohead',         1997),
        (None, 'Doggystyle',                        'Snoop Dogg',        1993),
        (None, 'The Slim Shady LP',                 'Eminem',            1999),
        (None, 'Ray of Light',                      'Madonna',           1998),
        (None, 'Jagged Little Pill',                'Alanis Morissette', 1995),
        (None, 'The College Dropout',               'Kanye West',        2004),
        (None, 'Graduation',                        'Kanye West',        2007),
        (None, 'good kid, m.A.A.d city',            'Kendrick Lamar',    2012),
        (None, 'To Pimp a Butterfly',               'Kendrick Lamar',    2015),
        (None, 'folklore',                          'Taylor Swift',      2020),
        (None, 'Random Access Memories',            'Daft Punk',         2013),
        (None, "Is This It",                        'The Strokes',       2001),
    ]

    for r, (rank, album, artist, year) in enumerate(user_albums, 2):
        ws.cell(row=r, column=1, value=rank)
        ws.cell(row=r, column=2, value=album)
        ws.cell(row=r, column=3, value=artist)
        ws.cell(row=r, column=4, value=year)

    # Column widths for readability
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 8

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup — open Chrome with Rolling Stone URL first, then LibreOffice Calc
    # Kill any existing Chrome instances to ensure a clean state
    subprocess.call(['pkill', '-f', 'chrome'], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    time.sleep(1.5)

    # Launch Chrome with the Rolling Stone 500 Greatest Albums page
    launch_gui(
        'google-chrome --remote-debugging-port=1337 '
        '"https://www.rollingstone.com/music/music-lists/best-albums-1291399/"',
        delay_sec=3.0
    )

    # Launch LibreOffice Calc with the album history file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)

    print('GUI_READY: launched Chrome (Rolling Stone URL) and LibreOffice Calc with DISPLAY=:0')


create_initial()
