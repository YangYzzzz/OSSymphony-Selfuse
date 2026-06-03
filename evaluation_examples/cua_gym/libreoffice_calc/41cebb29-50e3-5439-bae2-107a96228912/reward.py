"""
Reward Script: Rolling Stone Top 30 Must-Listen Albums
Task ID: osworld_multi_apps_misc_023
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): 'must_listen' sheet exists in my_albums.xlsx with correct headers
  Component 2 (0.4): Sheet contains correct albums from Rolling Stone top 30 not in
                     user's history, all released before 2020
  Component 3 (0.3): Albums are sorted by Rank ascending with correct rank values
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_misc_023'
FILE_PATH = f'{WORKDIR}/my_albums.xlsx'

# Expected headers for must_listen sheet (same as My Albums sheet)
EXPECTED_HEADERS = ['Rank', 'Album', 'Artist', 'Year']

# Known albums in user's "My Albums" history (by album title, normalized)
# These come from the initial state of My Albums sheet
MY_ALBUMS_TITLES = {
    'abbey road',
    'rumours',
    'the miseducation of lauryn hill',
    'thriller',
    'born to run',
    'led zeppelin iv',
    'highway 61 revisited',
    'kind of blue',
    'ok computer',
    'doggystyle',
    'the slim shady lp',
    'ray of light',
    'jagged little pill',
    'the college dropout',
    'graduation',
    'good kid, m.a.a.d city',
    'to pimp a butterfly',
    'folklore',
    'random access memories',
    'is this it',
}

# Ranks of albums already in user's listening history that are in top 30
MY_ALBUMS_RANKS_IN_TOP30 = {5, 7, 10, 15, 18, 21, 24, 30}


def has_contamination(data_rows):
    """Check if any album in data_rows is already in the user's listening history."""
    for row_data in data_rows:
        rank = row_data[0]
        album = row_data[1]
        if album and album.strip().lower() in MY_ALBUMS_TITLES:
            return album
        if rank is not None and rank in MY_ALBUMS_RANKS_IN_TOP30:
            return f"rank {rank}"
    return None


def has_post2019_album(data_rows):
    """Check if any album was released in 2020 or later."""
    for row_data in data_rows:
        album = row_data[1]
        year = row_data[3]
        if year is not None and isinstance(year, int) and year >= 2020:
            return (album, year)
    return None


def has_out_of_range_rank(data_rows):
    """Check if any album has rank outside 1-30."""
    for row_data in data_rows:
        album = row_data[1]
        rank = row_data[0]
        if rank is not None and isinstance(rank, int) and (rank < 1 or rank > 30):
            return (album, rank)
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: check that original "My Albums" sheet still exists
    if 'My Albums' not in wb.sheetnames:
        print("CRITICAL: 'My Albums' sheet is missing — file appears corrupted")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: 'must_listen' sheet exists with correct headers (0.3 points)
    # This FAILS on initial (no must_listen sheet) → PASSES on golden
    try:
        if 'must_listen' not in wb.sheetnames:
            print("FAIL: Component 1 — 'must_listen' sheet not found in workbook")
            print(f"  Found sheets: {wb.sheetnames}")
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score

        ws_ml = wb['must_listen']

        # Check headers match expected columns: Rank, Album, Artist, Year
        actual_headers = [ws_ml.cell(row=1, column=c).value for c in range(1, 5)]
        headers_match = (actual_headers == EXPECTED_HEADERS)
        if headers_match:
            print(f"PASS: Component 1 — 'must_listen' sheet exists with correct headers {actual_headers} (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — 'must_listen' sheet found but headers are wrong")
            print(f"  Expected: {EXPECTED_HEADERS}")
            print(f"  Found: {actual_headers}")

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Sheet contains correct albums — top-30 RS albums not already
    # in user's history and released before 2020 (0.4 points)
    # This FAILS on initial (no must_listen sheet) → PASSES on golden
    try:
        ws_ml = wb['must_listen']
        data_rows = []
        for row in ws_ml.iter_rows(min_row=2, max_row=ws_ml.max_row):
            vals = [c.value for c in row]
            if any(v is not None for v in vals):
                data_rows.append(vals)

        if len(data_rows) == 0:
            print("FAIL: Component 2 — 'must_listen' sheet has no data rows")
        else:
            contamination = has_contamination(data_rows)
            post2019 = has_post2019_album(data_rows)
            out_of_range = has_out_of_range_rank(data_rows)

            if contamination is None and post2019 is None and out_of_range is None and len(data_rows) >= 15:
                print(f"PASS: Component 2 — All {len(data_rows)} albums valid: top-30 RS not in user history, all pre-2020 (0.4 pts)")
                total_score += 0.4
            elif contamination is None and post2019 is None and out_of_range is None:
                print(f"PARTIAL: Component 2 — Albums valid but count low ({len(data_rows)}, expected ~22) (0.2 pts)")
                total_score += 0.2
            else:
                if contamination is not None:
                    print(f"FAIL: Component 2 — Found album already in user history: {contamination}")
                if post2019 is not None:
                    print(f"FAIL: Component 2 — Found post-2019 album: {post2019[0]} ({post2019[1]})")
                if out_of_range is not None:
                    print(f"FAIL: Component 2 — Found rank outside top 30: {out_of_range[0]} rank={out_of_range[1]}")

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Albums are sorted by Rank ascending (0.3 points)
    # This FAILS on initial (no must_listen sheet) → PASSES on golden
    try:
        ws_ml = wb['must_listen']
        ranks = []
        for row in ws_ml.iter_rows(min_row=2, max_row=ws_ml.max_row):
            rank_val = row[0].value
            if rank_val is not None:
                ranks.append(rank_val)

        if len(ranks) == 0:
            print("FAIL: Component 3 — No rank values found in must_listen sheet")
        else:
            is_sorted = all(ranks[i] <= ranks[i+1] for i in range(len(ranks)-1))
            all_valid_ranks = all(isinstance(r, int) and 1 <= r <= 30 for r in ranks)

            if is_sorted and all_valid_ranks:
                print(f"PASS: Component 3 — All {len(ranks)} albums sorted by Rank ascending, range {min(ranks)}-{max(ranks)} (0.3 pts)")
                total_score += 0.3
            elif is_sorted:
                print(f"PARTIAL: Component 3 — Sorted ascending but some rank values may be invalid (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 3 — Albums are NOT sorted by Rank ascending")
                print(f"  Ranks found: {ranks}")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Main entry point
if not os.path.exists(FILE_PATH):
    print(f"File not found: {FILE_PATH}")
    print("REWARD: 0.0")
else:
    verify_task(FILE_PATH)
