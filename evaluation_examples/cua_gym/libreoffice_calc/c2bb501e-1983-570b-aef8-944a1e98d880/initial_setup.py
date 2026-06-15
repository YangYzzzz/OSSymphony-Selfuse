"""
Initial Setup: Deal scoring model with weighted criteria
Task ID: calc_sales_057
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_057'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Scoring ---
    ws = wb.active
    ws.title = 'Scoring'

    # Headers
    headers = [
        'Deal', 'Size Score (1-5)', 'Stage Score (1-5)',
        'Engagement (1-5)', 'Timeline (1-5)', 'Champion (1-5)',
        'Composite Score', 'Priority'
    ]
    header_font = Font(bold=True, size=11, name='Calibri')
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, name='Calibri', color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows (B-F only; G and H left empty for the task)
    data = [
        ['Deal Alpha',   5, 4, 3, 4, 5],
        ['Deal Beta',    3, 2, 4, 3, 2],
        ['Deal Gamma',   4, 5, 5, 5, 4],
        ['Deal Delta',   2, 3, 2, 2, 1],
        ['Deal Epsilon', 4, 4, 4, 3, 3],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 1:
                cell.font = Font(bold=True, name='Calibri')
            else:
                cell.alignment = Alignment(horizontal="center")

    # G and H columns left EMPTY (task is to fill these)

    # Weights reference in column I-J
    ws.cell(row=1, column=9, value='Criterion').font = Font(bold=True, name='Calibri')
    ws.cell(row=1, column=10, value='Weight').font = Font(bold=True, name='Calibri')

    weight_labels = ['Size Score', 'Stage Score', 'Engagement', 'Timeline', 'Champion']
    weight_values = [0.30, 0.25, 0.20, 0.15, 0.10]

    for i, (label, weight) in enumerate(zip(weight_labels, weight_values), 2):
        ws.cell(row=i, column=9, value=label)
        cell = ws.cell(row=i, column=10, value=weight)
        cell.number_format = '0%'
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 10

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
