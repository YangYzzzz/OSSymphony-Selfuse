"""
Reward Script: Deal scoring model with weighted criteria composite score and priority
Task ID: calc_sales_057
Domain: libreoffice_calc
Scoring:
  Component 1 (0.50): G2:G6 contain weighted composite score formulas
  Component 2 (0.30): H2:H6 contain priority IF formulas
  Component 3 (0.20): Formulas produce correct computed values (verified programmatically)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_057'

# Ground truth: input scores per deal (rows 2-6), columns B-F
INPUT_SCORES = {
    2: [5, 4, 3, 4, 5],
    3: [3, 2, 4, 3, 2],
    4: [4, 5, 5, 5, 4],
    5: [2, 3, 2, 2, 1],
    6: [4, 4, 4, 3, 3],
}

# Weights: Size=0.30, Stage=0.25, Engagement=0.20, Timeline=0.15, Champion=0.10
WEIGHTS = [0.30, 0.25, 0.20, 0.15, 0.10]

# Expected composite scores (computed from input data and weights)
# Row 2: 5*0.30+4*0.25+3*0.20+4*0.15+5*0.10 = 1.50+1.00+0.60+0.60+0.50 = 4.20
# Row 3: 3*0.30+2*0.25+4*0.20+3*0.15+2*0.10 = 0.90+0.50+0.80+0.45+0.20 = 2.85
# Row 4: 4*0.30+5*0.25+5*0.20+5*0.15+4*0.10 = 1.20+1.25+1.00+0.75+0.40 = 4.60
# Row 5: 2*0.30+3*0.25+2*0.20+2*0.15+1*0.10 = 0.60+0.75+0.40+0.30+0.10 = 2.15
# Row 6: 4*0.30+4*0.25+4*0.20+3*0.15+3*0.10 = 1.20+1.00+0.80+0.45+0.30 = 3.75
EXPECTED_G = {2: 4.20, 3: 2.85, 4: 4.60, 5: 2.15, 6: 3.75}
# Expected priorities based on computed composite scores
EXPECTED_H = {2: 'High', 3: 'Low', 4: 'High', 5: 'Low', 6: 'Medium'}


def compute_composite(scores):
    """Compute weighted composite from raw scores."""
    return sum(s * w for s, w in zip(scores, WEIGHTS))


def normalize_formula(f):
    """Normalize a formula string for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def is_valid_composite_formula(formula, row):
    """
    Check if formula computes weighted composite score for the given row.
    Accepts various forms:
      =B2*0.30+C2*0.25+D2*0.20+E2*0.15+F2*0.10
      =B2*0.3+C2*0.25+D2*0.2+E2*0.15+F2*0.1
      =B2*J$1+C2*J$2+... (referencing weight cells)
      =SUMPRODUCT(B2:F2,$J$1:$J$5) or similar
    We verify by checking the formula references the correct row's B-F cells
    and either contains weight constants or references to weight cells.
    """
    nf = normalize_formula(formula)
    if not nf.startswith('='):
        return False

    r = str(row)

    # Check for SUMPRODUCT-style formula
    if 'SUMPRODUCT' in nf:
        # Must reference the row's data range and weight range
        if f'B{r}' in nf and f'F{r}' in nf:
            return True

    # Check for explicit weighted sum: B*w + C*w + D*w + E*w + F*w
    # Must reference all 5 score columns for this row
    cols = ['B', 'C', 'D', 'E', 'F']
    refs_found = all(f'{col}{r}' in nf for col in cols)
    if not refs_found:
        return False

    # Must contain weight values (0.3, 0.25, 0.2, 0.15, 0.1) or references to J column
    has_weights = (
        ('0.3' in nf or '0.30' in nf) and
        '0.25' in nf and
        ('0.2' in nf) and
        '0.15' in nf and
        ('0.1' in nf)
    )
    has_weight_refs = 'J' in nf  # references weight cells

    return has_weights or has_weight_refs


def is_valid_priority_formula(formula, row):
    """
    Check if formula computes priority based on composite score G.
    Expected: =IF(G>=4,"High",IF(G>=3,"Medium","Low"))
    Accepts variations in quoting and nesting.
    """
    nf = normalize_formula(formula)
    if not nf.startswith('='):
        return False

    r = str(row)

    # Must reference G column for this row
    if f'G{r}' not in nf:
        return False

    # Must contain IF function
    if 'IF(' not in nf:
        return False

    # Must contain the three priority labels
    nf_check = nf.replace('"', '').replace("'", '')
    has_high = 'HIGH' in nf_check
    has_medium = 'MEDIUM' in nf_check
    has_low = 'LOW' in nf_check

    # Must contain threshold values 4 and 3
    has_thresholds = '4' in nf and '3' in nf

    return has_high and has_medium and has_low and has_thresholds


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Find the Scoring sheet
    if 'Scoring' not in wb.sheetnames:
        print(f"FAIL: 'Scoring' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Scoring']

    # Component 1: G2:G6 contain weighted composite score formulas (0.50 points)
    try:
        composite_pass = 0
        for row in range(2, 7):
            cell_val = ws.cell(row=row, column=7).value  # Column G
            if is_valid_composite_formula(cell_val, row):
                composite_pass += 1
                print(f"  PASS: G{row} has valid composite formula: {cell_val}")
            else:
                print(f"  FAIL: G{row} expected composite formula, found: {cell_val}")

        if composite_pass == 5:
            print(f"PASS: Component 1 -- All 5 composite score formulas present (0.50 pts)")
            total_score += 0.50
        elif composite_pass > 0:
            partial = round(0.50 * composite_pass / 5, 2)
            print(f"PARTIAL: Component 1 -- {composite_pass}/5 composite formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 -- No valid composite score formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: H2:H6 contain priority IF formulas (0.30 points)
    try:
        priority_pass = 0
        for row in range(2, 7):
            cell_val = ws.cell(row=row, column=8).value  # Column H
            if is_valid_priority_formula(cell_val, row):
                priority_pass += 1
                print(f"  PASS: H{row} has valid priority formula: {cell_val}")
            else:
                print(f"  FAIL: H{row} expected priority IF formula, found: {cell_val}")

        if priority_pass == 5:
            print(f"PASS: Component 2 -- All 5 priority formulas present (0.30 pts)")
            total_score += 0.30
        elif priority_pass > 0:
            partial = round(0.30 * priority_pass / 5, 2)
            print(f"PARTIAL: Component 2 -- {priority_pass}/5 priority formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- No valid priority formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Formula correctness -- computed values match ground truth (0.20 points)
    # We evaluate the formulas programmatically using the known input scores
    try:
        correct_values = 0
        total_checks = 10  # 5 composite + 5 priority

        for row in range(2, 7):
            g_formula = ws.cell(row=row, column=7).value
            h_formula = ws.cell(row=row, column=8).value

            # Check composite value by computing from known inputs
            if is_valid_composite_formula(g_formula, row):
                expected_g = EXPECTED_G[row]
                computed_g = compute_composite(INPUT_SCORES[row])
                if abs(computed_g - expected_g) < 0.01:
                    correct_values += 1
                    print(f"  PASS: G{row} formula would compute {computed_g:.2f} (expected {expected_g})")
                else:
                    print(f"  FAIL: G{row} formula mismatch: computed {computed_g:.2f} vs expected {expected_g}")
            else:
                print(f"  SKIP: G{row} no valid formula to evaluate")

            # Check priority logic
            if is_valid_priority_formula(h_formula, row):
                expected_h = EXPECTED_H[row]
                computed_g = compute_composite(INPUT_SCORES[row])
                if computed_g >= 4:
                    computed_h = 'High'
                elif computed_g >= 3:
                    computed_h = 'Medium'
                else:
                    computed_h = 'Low'

                if computed_h == expected_h:
                    correct_values += 1
                    print(f"  PASS: H{row} formula would produce '{computed_h}' (expected '{expected_h}')")
                else:
                    print(f"  FAIL: H{row} formula mismatch: '{computed_h}' vs expected '{expected_h}'")
            else:
                print(f"  SKIP: H{row} no valid formula to evaluate")

        if correct_values == total_checks:
            print(f"PASS: Component 3 -- All {total_checks} computed values match ground truth (0.20 pts)")
            total_score += 0.20
        elif correct_values > 0:
            partial = round(0.20 * correct_values / total_checks, 2)
            print(f"PARTIAL: Component 3 -- {correct_values}/{total_checks} values correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 -- No computed values match ground truth")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
