"""
Reward Script: Protect 'Template' sheet with password and specific allowed operations
Task ID: calc_ps_011
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Sheet protection is enabled
  Component 2 (0.20): Password hash is set
  Component 3 (0.15): Insert columns is allowed (insertColumns=False when protected)
  Component 4 (0.15): Delete columns is allowed (deleteColumns=False when protected)
  Component 5 (0.20): Delete rows is allowed (deleteRows=False when protected)
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_011'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    In openpyxl SheetProtection, when sheet=True (protection enabled):
    - A flag like insertColumns=False means the action IS ALLOWED
    - A flag like insertColumns=True means the action IS FORBIDDEN
    This is because the flags represent "protect against this action".
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Template' sheet must exist
    if 'Template' not in wb.sheetnames:
        print(f"CRITICAL: 'Template' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Template']
    prot = ws.protection

    # Component 1: Sheet protection is enabled (0.30 points)
    try:
        if prot.sheet is True:
            print(f"PASS: Component 1 — Sheet protection enabled (sheet={prot.sheet}) (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 — Sheet protection not enabled (sheet={prot.sheet})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Password hash is set (0.20 points)
    # When protected with a password, openpyxl stores a hash value
    try:
        has_password = (prot.password is not None and prot.password != '')
        if has_password:
            print(f"PASS: Component 2 — Password hash is set (hash={prot.password}) (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 — No password set (password={repr(prot.password)})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Insert columns is allowed (0.15 points)
    # When sheet is protected, insertColumns=False means the action is ALLOWED
    try:
        if prot.sheet and prot.insertColumns is False:
            print(f"PASS: Component 3 — Insert columns allowed (insertColumns={prot.insertColumns}) (0.15 pts)")
            total_score += 0.15
        else:
            if not prot.sheet:
                print(f"FAIL: Component 3 — Sheet not protected, cannot verify insertColumns permission")
            else:
                print(f"FAIL: Component 3 — Insert columns not allowed (insertColumns={prot.insertColumns})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Delete columns is allowed (0.15 points)
    try:
        if prot.sheet and prot.deleteColumns is False:
            print(f"PASS: Component 4 — Delete columns allowed (deleteColumns={prot.deleteColumns}) (0.15 pts)")
            total_score += 0.15
        else:
            if not prot.sheet:
                print(f"FAIL: Component 4 — Sheet not protected, cannot verify deleteColumns permission")
            else:
                print(f"FAIL: Component 4 — Delete columns not allowed (deleteColumns={prot.deleteColumns})")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Delete rows is allowed (0.20 points)
    try:
        if prot.sheet and prot.deleteRows is False:
            print(f"PASS: Component 5 — Delete rows allowed (deleteRows={prot.deleteRows}) (0.20 pts)")
            total_score += 0.20
        else:
            if not prot.sheet:
                print(f"FAIL: Component 5 — Sheet not protected, cannot verify deleteRows permission")
            else:
                print(f"FAIL: Component 5 — Delete rows not allowed (deleteRows={prot.deleteRows})")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist app state before verification
persist_app_state("libreoffice_calc")

# Run verification
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
