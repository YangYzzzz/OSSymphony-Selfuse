"""
Initial Setup: Create a spreadsheet with an unprotected 'Template' sheet containing a data table.
Task ID: calc_ps_011
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_011'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Template'

    # Headers
    headers = ['Employee', 'Department', 'Quarterly Revenue', 'Join Date']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows (19 rows of realistic employee data)
    data = [
        ['Sarah Chen', 'Engineering', 87500, '2022-03-15'],
        ['Marcus Johnson', 'Marketing', 64200, '2021-08-01'],
        ['Priya Patel', 'Engineering', 92300, '2020-11-20'],
        ['David Kim', 'Finance', 78100, '2023-01-10'],
        ['Elena Rodriguez', 'Sales', 71800, '2022-06-25'],
        ['James Williams', 'Engineering', 95400, '2019-04-12'],
        ['Aisha Mohammed', 'Marketing', 61500, '2023-05-03'],
        ['Robert Taylor', 'Finance', 83700, '2021-02-18'],
        ['Lisa Wang', 'Sales', 76900, '2022-09-30'],
        ['Michael Brown', 'Engineering', 89200, '2020-07-14'],
        ['Jennifer Lopez', 'HR', 58300, '2023-03-22'],
        ['Thomas Anderson', 'Sales', 74600, '2021-12-05'],
        ['Sophia Nguyen', 'Finance', 81400, '2022-01-28'],
        ['Daniel Martinez', 'Engineering', 93100, '2019-10-17'],
        ['Rachel Green', 'Marketing', 67800, '2023-07-09'],
        ['Kevin O\'Brien', 'HR', 55900, '2022-04-11'],
        ['Maria Santos', 'Sales', 72400, '2021-06-30'],
        ['Andrew Clark', 'Finance', 80200, '2020-09-08'],
        ['Hannah Foster', 'Engineering', 91700, '2023-02-14'],
    ]

    data_font = Font(name='Calibri', size=11)
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = thin_border
            if c == 3:  # Revenue column
                cell.number_format = '$#,##0'
            elif c == 4:  # Date column
                cell.alignment = Alignment(horizontal='center')

    # Set column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
