"""
Initial Setup: Fill geometric series in RevenueProjection sheet
Task ID: calc_dop_fillseries_geometric_049
Domain: libreoffice_calc

Creates an initial spreadsheet with RevenueProjection sheet containing:
- Month numbers (A2:A13), Month names (B2:B13)
- C2 = 1000 (starting value already filled in)
- C3:C13 empty (agent must fill using geometric series)
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_dop_fillseries_geometric_049'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: RevenueProjection ---
    ws = wb.active
    ws.title = 'RevenueProjection'

    # Headers in row 1
    ws['A1'] = 'Month'
    ws['B1'] = 'Month Name'
    ws['C1'] = 'Projected Revenue'

    # Month numbers and names
    month_names = [
        'January', 'February', 'March', 'April',
        'May', 'June', 'July', 'August',
        'September', 'October', 'November', 'December'
    ]

    for i, name in enumerate(month_names, start=1):
        row = i + 1  # data starts at row 2
        ws.cell(row=row, column=1, value=i)       # Month number
        ws.cell(row=row, column=2, value=name)    # Month name
        # C column: only C2 has the starting value 1000; C3:C13 are empty
        if i == 1:
            ws.cell(row=row, column=3, value=1000)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('  Sheet: RevenueProjection')
    print('  A2:A13: months 1-12')
    print('  B2:B13: January through December')
    print('  C2: 1000 (starting value)')
    print('  C3:C13: empty (to be filled by agent)')

create_initial()
