"""
Reward Script: Fill C2:C13 with a geometric growth series starting at 1000, multiplying by 1.1 each month.
Task ID: calc_dop_fillseries_geometric_049
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.4): C3:C13 are all non-empty (series was actually filled)
  - Component 2 (0.4): C3:C13 follow a geometric progression with factor ~1.1
  - Component 3 (0.2): C2 = 1000 (starting value intact) and C13 ≈ 2853.12 (end value correct)
"""

import os
import math

import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_dop_fillseries_geometric_049'
SHEET_NAME = 'RevenueProjection'

# Expected geometric series values: C2=1000, Cn = 1000 * 1.1^(n-2)
GROWTH_FACTOR = 1.1
BASE_VALUE = 1000.0
# Values for C2..C13 (rows 2..13, 12 entries)
EXPECTED_VALUES = [BASE_VALUE * (GROWTH_FACTOR ** i) for i in range(12)]


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet exists
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Read C2:C13 values (rows 2 through 13, column 3)
    c_values = []
    for row in range(2, 14):
        val = ws.cell(row=row, column=3).value
        c_values.append(val)

    # c_values[0] = C2, c_values[1] = C3, ..., c_values[11] = C13

    # Component 1: C3:C13 are all non-empty (cells C3..C13 were filled, indices 1..11)
    # The task asks to fill C2:C13; C2 was already 1000, so the change is C3:C13 being filled.
    try:
        cells_to_check = c_values[1:]  # C3:C13 (indices 1..11)
        non_empty_count = sum(1 for v in cells_to_check if v is not None)
        if non_empty_count == 11:
            print(f"PASS: Component 1 — All 11 cells C3:C13 are non-empty (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — Only {non_empty_count}/11 cells in C3:C13 are non-empty")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: C3:C13 follow a geometric progression with growth factor ~1.1
    # Each consecutive pair should have a ratio of approximately 1.1
    # We check all 11 values in C3:C13 against expected values derived from 1000 * 1.1^k
    try:
        tolerance = 0.01  # 1% relative tolerance
        geometric_correct = 0
        geometric_total = 0

        for idx in range(1, 12):  # indices 1..11 correspond to C3..C13
            row_num = idx + 2  # row number (C3 = row 3)
            val = c_values[idx]
            expected = EXPECTED_VALUES[idx]  # 1000 * 1.1^idx
            geometric_total += 1
            if val is not None:
                try:
                    numeric_val = float(val)
                    rel_error = abs(numeric_val - expected) / expected
                    if rel_error <= tolerance:
                        geometric_correct += 1
                    else:
                        print(f"  FAIL: C{row_num} = {numeric_val}, expected ~{expected:.4f} (rel_error={rel_error:.4f})")
                except (ValueError, TypeError):
                    print(f"  FAIL: C{row_num} value '{val}' is not numeric")
            else:
                print(f"  FAIL: C{row_num} is None/empty")

        if geometric_correct == 11:
            print(f"PASS: Component 2 — All 11 values in C3:C13 match geometric series (factor=1.1) (0.4 pts)")
            total_score += 0.4
        elif geometric_correct >= 6:
            print(f"PARTIAL: Component 2 — {geometric_correct}/11 values match geometric series. Awarding partial (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 2 — Only {geometric_correct}/11 values match the geometric series")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: C2 = 1000 (starting value intact) AND C13 ≈ 2853.12 (final value correct)
    # Both must hold for full credit (0.2 pts)
    try:
        c2_val = c_values[0]  # C2
        c13_val = c_values[11]  # C13

        c2_ok = False
        c13_ok = False

        if c2_val is not None:
            try:
                c2_ok = abs(float(c2_val) - 1000.0) <= 0.01
            except (ValueError, TypeError):
                pass

        if c13_val is not None:
            try:
                # Expected C13 = 1000 * 1.1^11 = 2853.11670611
                expected_c13 = 1000.0 * (1.1 ** 11)
                c13_ok = abs(float(c13_val) - expected_c13) / expected_c13 <= 0.01
            except (ValueError, TypeError):
                pass

        if c2_ok and c13_ok:
            print(f"PASS: Component 3 — C2={c2_val} (expected 1000) and C13={c13_val} (expected ~2853.12) (0.2 pts)")
            total_score += 0.2
        elif c2_ok:
            print(f"FAIL: Component 3 — C2={c2_val} OK, but C13={c13_val} not correct (expected ~2853.12)")
        elif c13_ok:
            print(f"FAIL: Component 3 — C13={c13_val} OK, but C2={c2_val} not equal to 1000")
        else:
            print(f"FAIL: Component 3 — C2={c2_val} (expected 1000), C13={c13_val} (expected ~2853.12)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
