"""
Reward Script: Cross-sheet conditional sum formulas on Pivot sheet
Task ID: calc_mcp_060
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): All 8 cells B2:C5 contain SUMIFS formulas
  Component 2 (0.3): Formulas reference correct Transactions sheet columns
  Component 3 (0.3): Formulas use proper mixed references for drag-fill pattern
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_060'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Pivot sheet must exist
    if 'Pivot' not in wb.sheetnames:
        print("FAIL: 'Pivot' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Pivot']

    # Collect all formulas from B2:C5
    target_cells = []
    for row in range(2, 6):
        for col_letter in ['B', 'C']:
            coord = f"{col_letter}{row}"
            val = ws[coord].value
            target_cells.append((coord, val))

    # Component 1: All 8 cells B2:C5 contain SUMIFS formulas (0.4 points)
    # This checks that the task was done at all - cells should have formulas, not be empty
    try:
        sumifs_count = 0
        for coord, val in target_cells:
            if val is not None and isinstance(val, str) and 'SUMIFS' in val.upper():
                sumifs_count += 1

        if sumifs_count == 8:
            print(f"PASS: Component 1 - All 8 cells contain SUMIFS formulas (0.4 pts)")
            total_score += 0.4
        elif sumifs_count > 0:
            partial = round(0.4 * (sumifs_count / 8), 2)
            print(f"PARTIAL: Component 1 - {sumifs_count}/8 cells have SUMIFS formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - No SUMIFS formulas found in B2:C5. Values: {[v for _, v in target_cells]}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Formulas reference correct Transactions sheet and columns (0.3 points)
    # The formulas must reference Transactions!E (sum range), Transactions!B (dept criteria),
    # and Transactions!C (quarter criteria)
    try:
        correct_refs_count = 0
        for coord, val in target_cells:
            if val is not None and isinstance(val, str):
                formula_upper = val.upper().replace(" ", "")
                # Check for references to Transactions sheet columns E, B, and C
                has_sum_range = bool(re.search(r"TRANSACTIONS!E", formula_upper))
                has_dept_criteria = bool(re.search(r"TRANSACTIONS!B", formula_upper))
                has_quarter_criteria = bool(re.search(r"TRANSACTIONS!C", formula_upper))
                if has_sum_range and has_dept_criteria and has_quarter_criteria:
                    correct_refs_count += 1

        if correct_refs_count == 8:
            print(f"PASS: Component 2 - All 8 formulas reference correct Transactions columns (0.3 pts)")
            total_score += 0.3
        elif correct_refs_count > 0:
            partial = round(0.3 * (correct_refs_count / 8), 2)
            print(f"PARTIAL: Component 2 - {correct_refs_count}/8 formulas have correct refs ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 - No formulas correctly reference Transactions!E, B, C columns")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Formulas use correct cell references for department (column A) and
    # quarter header (row 1) so the formula pattern works across the grid (0.3 points)
    # Expected pattern: criteria for dept references $A<row> (locked column, variable row)
    # and criteria for quarter references <col>$1 (variable column, locked row)
    try:
        correct_pattern_count = 0
        for coord, val in target_cells:
            if val is not None and isinstance(val, str):
                formula_norm = val.replace(" ", "")
                # Check for department reference: $A followed by the row number of this cell
                row_num = coord[1]  # e.g., '2' from 'B2'
                col_letter = coord[0]  # e.g., 'B' from 'B2'

                # The dept criteria should reference A-column for the same row
                # Could be $A2, $A$2, A2, etc. - key is it references A and the current row
                has_dept_ref = bool(re.search(
                    r'[\$]?A[\$]?' + row_num, formula_norm
                ))

                # The quarter criteria should reference row 1 with the same column
                # Could be B$1, $B$1, B1, etc. - key is it references the column and row 1
                has_quarter_ref = bool(re.search(
                    col_letter + r'[\$]?1', formula_norm
                ))

                if has_dept_ref and has_quarter_ref:
                    correct_pattern_count += 1

        if correct_pattern_count == 8:
            print(f"PASS: Component 3 - All 8 formulas use correct cell references for dept/quarter (0.3 pts)")
            total_score += 0.3
        elif correct_pattern_count > 0:
            partial = round(0.3 * (correct_pattern_count / 8), 2)
            print(f"PARTIAL: Component 3 - {correct_pattern_count}/8 formulas have correct references ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 - No formulas have correct dept/quarter cell references")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
