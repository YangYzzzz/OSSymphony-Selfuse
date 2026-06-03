"""
Initial Setup: Cross-sheet conditional sum formulas on Pivot sheet
Task ID: calc_mcp_060
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_060'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Transactions ---
    ws_txn = wb.active
    ws_txn.title = 'Transactions'

    headers = ['ID', 'Dept', 'Quarter', 'Type', 'Amount']
    for col, h in enumerate(headers, 1):
        ws_txn.cell(row=1, column=col, value=h)

    # Realistic transaction data across 4 departments and 2 quarters
    transactions = [
        ['TXN-001', 'Engineering',  'Q1', 'Software License',  4500.00],
        ['TXN-002', 'Marketing',    'Q1', 'Ad Campaign',       12300.00],
        ['TXN-003', 'Engineering',  'Q2', 'Cloud Hosting',     8750.00],
        ['TXN-004', 'Sales',        'Q1', 'Travel Expense',    3200.00],
        ['TXN-005', 'HR',           'Q1', 'Recruiting Fee',    6800.00],
        ['TXN-006', 'Marketing',    'Q2', 'Event Sponsorship', 9500.00],
        ['TXN-007', 'Engineering',  'Q1', 'Hardware Purchase',  7200.00],
        ['TXN-008', 'Sales',        'Q2', 'Client Dinner',     1850.00],
        ['TXN-009', 'HR',           'Q2', 'Training Program',  4300.00],
        ['TXN-010', 'Engineering',  'Q2', 'Conference Ticket',  2100.00],
        ['TXN-011', 'Marketing',    'Q1', 'Print Materials',   3400.00],
        ['TXN-012', 'Sales',        'Q1', 'Commission Payout', 5600.00],
        ['TXN-013', 'HR',           'Q1', 'Office Supplies',   1250.00],
        ['TXN-014', 'Sales',        'Q2', 'Software Demo',     2900.00],
        ['TXN-015', 'Marketing',    'Q2', 'Social Media Ads',  7800.00],
        ['TXN-016', 'Engineering',  'Q1', 'Dev Tools',         3150.00],
        ['TXN-017', 'HR',           'Q2', 'Team Building',     2750.00],
        ['TXN-018', 'Sales',        'Q2', 'Trade Show Booth',  6100.00],
    ]

    for r, row_data in enumerate(transactions, 2):
        for c, val in enumerate(row_data, 1):
            ws_txn.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws_txn.column_dimensions['A'].width = 12
    ws_txn.column_dimensions['B'].width = 15
    ws_txn.column_dimensions['C'].width = 10
    ws_txn.column_dimensions['D'].width = 22
    ws_txn.column_dimensions['E'].width = 14

    # --- Sheet 2: Pivot ---
    ws_pivot = wb.create_sheet('Pivot')
    ws_pivot['A1'] = 'Department'
    ws_pivot['B1'] = 'Q1'
    ws_pivot['C1'] = 'Q2'

    departments = ['Engineering', 'Marketing', 'Sales', 'HR']
    for i, dept in enumerate(departments, 2):
        ws_pivot.cell(row=i, column=1, value=dept)

    # B2:C5 intentionally left EMPTY — task is to fill them with SUMIFS formulas
    ws_pivot.column_dimensions['A'].width = 15
    ws_pivot.column_dimensions['B'].width = 12
    ws_pivot.column_dimensions['C'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
