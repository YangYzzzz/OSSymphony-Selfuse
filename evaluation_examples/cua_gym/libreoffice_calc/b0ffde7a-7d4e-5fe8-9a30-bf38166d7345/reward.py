"""
Reward Script: ETL Pipeline + LibreOffice Calc Pivot Summary
Task ID: osworld_multi_apps_code_batch_terminal_012
Domain: multi_apps (bash, python, libreoffice_calc)
Scoring:
  Component 1: ETL pipeline script exists at /home/user/scripts/etl_pipeline.sh (0.20 pts)
  Component 2: 5 cleaned CSV files in /home/user/data/clean/ with YYYY-MM-DD dates, no missing values (0.30 pts)
  Component 3: combined.csv at /home/user/data/final/ with correct header, deduplicated rows, valid dates (0.20 pts)
  Component 4: xlsx Sheet2 exists with region-based sales totals matching combined.csv (0.30 pts)
  Total: 1.0
"""

import os
import csv
import re

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_code_batch_terminal_012'


def verify_task():
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # -----------------------------------------------------------------------
    # Component 1: ETL pipeline script exists (0.20 points)
    # The task requires writing /home/user/scripts/etl_pipeline.sh
    # This FAILS on initial (no script) and PASSES on golden (script exists)
    # -----------------------------------------------------------------------
    try:
        script_path = f'{WORKDIR}/scripts/etl_pipeline.sh'
        if os.path.isfile(script_path):
            # Also verify it has some executable content (not empty)
            with open(script_path, 'r') as f:
                content = f.read()
            if len(content.strip()) > 20:
                print(f"PASS: Component 1 — etl_pipeline.sh exists and has content (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 1 — etl_pipeline.sh exists but is nearly empty")
        else:
            print(f"FAIL: Component 1 — etl_pipeline.sh not found at {script_path}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: 5 cleaned CSV files in /home/user/data/clean/ (0.30 points)
    # Each file must:
    #   - Have header: date,product,sales,region
    #   - Use YYYY-MM-DD date format (normalized)
    #   - Have no rows with missing values
    # This FAILS on initial (clean/ is empty) and PASSES on golden (5 clean CSVs)
    # -----------------------------------------------------------------------
    try:
        clean_dir = f'{WORKDIR}/data/clean'
        if not os.path.isdir(clean_dir):
            print(f"FAIL: Component 2 — clean directory does not exist")
        else:
            clean_files = [f for f in os.listdir(clean_dir) if f.endswith('.csv')]
            if len(clean_files) < 5:
                print(f"FAIL: Component 2 — only {len(clean_files)} CSV files in clean/ (need 5)")
            else:
                date_pattern = re.compile(r'^\d{4}-\d{2}-\d{2}$')
                expected_header = ['date', 'product', 'sales', 'region']
                all_ok = True
                issues = []
                for fname in clean_files:
                    fpath = os.path.join(clean_dir, fname)
                    try:
                        with open(fpath, 'r') as f:
                            reader = csv.DictReader(f)
                            # Check header
                            if reader.fieldnames != expected_header:
                                issues.append(f"{fname}: wrong header {reader.fieldnames}")
                                all_ok = False
                                continue
                            rows = list(reader)
                        # Check for missing values
                        missing_rows = [r for r in rows if not r.get('date') or not r.get('product') or not r.get('sales') or not r.get('region')]
                        if missing_rows:
                            issues.append(f"{fname}: {len(missing_rows)} rows with missing values")
                            all_ok = False
                        # Check date format
                        bad_dates = [r for r in rows if not date_pattern.match(r.get('date', ''))]
                        if bad_dates:
                            issues.append(f"{fname}: {len(bad_dates)} rows with non-YYYY-MM-DD dates")
                            all_ok = False
                    except Exception as fe:
                        issues.append(f"{fname}: error reading — {fe}")
                        all_ok = False

                if all_ok:
                    print(f"PASS: Component 2 — {len(clean_files)} cleaned CSVs with valid headers, dates, and no missing values (0.30 pts)")
                    total_score += 0.30
                else:
                    print(f"FAIL: Component 2 — issues in cleaned CSVs: {'; '.join(issues)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: combined.csv at /home/user/data/final/ (0.20 points)
    # Must have: header date,product,sales,region, >=40 data rows (5 files x 8+rows),
    # all YYYY-MM-DD dates, no missing values, and deduplicated (date+product unique)
    # This FAILS on initial (final/ is empty) and PASSES on golden
    # -----------------------------------------------------------------------
    try:
        combined_path = f'{WORKDIR}/data/final/combined.csv'
        if not os.path.isfile(combined_path):
            print(f"FAIL: Component 3 — combined.csv not found at {combined_path}")
        else:
            with open(combined_path, 'r') as f:
                reader = csv.DictReader(f)
                expected_header = ['date', 'product', 'sales', 'region']
                if reader.fieldnames != expected_header:
                    print(f"FAIL: Component 3 — combined.csv has wrong header: {reader.fieldnames}")
                else:
                    rows = list(reader)
                    if len(rows) < 40:
                        print(f"FAIL: Component 3 — combined.csv has only {len(rows)} rows (expected >= 40)")
                    else:
                        date_pattern = re.compile(r'^\d{4}-\d{2}-\d{2}$')
                        # Check no missing values
                        missing_rows = [r for r in rows if not r.get('date') or not r.get('product') or not r.get('sales') or not r.get('region')]
                        # Check date format
                        bad_dates = [r for r in rows if not date_pattern.match(r.get('date', ''))]
                        # Check deduplication (date+product unique)
                        combos = [(r['date'], r['product']) for r in rows]
                        unique_combos = set(combos)
                        is_deduplicated = len(unique_combos) == len(combos)

                        if missing_rows:
                            print(f"FAIL: Component 3 — combined.csv has {len(missing_rows)} rows with missing values")
                        elif bad_dates:
                            print(f"FAIL: Component 3 — combined.csv has {len(bad_dates)} rows with non-YYYY-MM-DD dates")
                        elif not is_deduplicated:
                            dupes = len(combos) - len(unique_combos)
                            print(f"FAIL: Component 3 — combined.csv has {dupes} duplicate date+product combos")
                        else:
                            print(f"PASS: Component 3 — combined.csv has {len(rows)} deduplicated rows, correct header, YYYY-MM-DD dates (0.20 pts)")
                            total_score += 0.20
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: xlsx Sheet2 has region-based sales summary (0.30 points)
    # Sheet2 must exist with a Region/Total Sales table that matches combined.csv
    # This FAILS on initial (no xlsx file) and PASSES on golden
    # -----------------------------------------------------------------------
    try:
        import openpyxl
        xlsx_path = f'{WORKDIR}/{TASK_ID}.xlsx'
        if not os.path.isfile(xlsx_path):
            print(f"FAIL: Component 4 — xlsx file not found at {xlsx_path}")
        else:
            wb = openpyxl.load_workbook(xlsx_path)
            if 'Sheet2' not in wb.sheetnames:
                print(f"FAIL: Component 4 — Sheet2 not found in {xlsx_path}. Sheets: {wb.sheetnames}")
            else:
                ws2 = wb['Sheet2']
                # Collect all non-empty rows from Sheet2
                sheet2_data = {}
                regions_found = []
                for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, values_only=True):
                    if row[0] is not None and row[1] is not None:
                        key = str(row[0]).strip()
                        val = row[1]
                        if key.lower() not in ('region', 'total'):
                            sheet2_data[key] = val

                # Compute expected region totals from combined.csv (if it exists)
                combined_path = f'{WORKDIR}/data/final/combined.csv'
                expected_regions = {}
                if os.path.isfile(combined_path):
                    with open(combined_path, 'r') as f:
                        reader = csv.DictReader(f)
                        for r in reader:
                            if r.get('region') and r.get('sales'):
                                try:
                                    expected_regions[r['region']] = expected_regions.get(r['region'], 0.0) + float(r['sales'])
                                except ValueError:
                                    pass

                if len(sheet2_data) < 2:
                    print(f"FAIL: Component 4 — Sheet2 has fewer than 2 region rows: {sheet2_data}")
                elif len(expected_regions) > 0:
                    # Verify at least 3 regions match within tolerance
                    matched = 0
                    for region, expected_total in expected_regions.items():
                        if region in sheet2_data:
                            try:
                                actual = float(sheet2_data[region])
                                if abs(actual - expected_total) <= 1.0:
                                    matched += 1
                            except (TypeError, ValueError):
                                pass

                    if matched >= len(expected_regions):
                        print(f"PASS: Component 4 — Sheet2 has {len(sheet2_data)} regions with correct totals matching combined.csv (0.30 pts)")
                        total_score += 0.30
                    elif matched >= 3:
                        # Partial — at least 3 regions correct
                        partial = 0.15
                        print(f"PARTIAL: Component 4 — {matched}/{len(expected_regions)} region totals match (0.15 pts)")
                        total_score += partial
                    else:
                        print(f"FAIL: Component 4 — only {matched}/{len(expected_regions)} region totals match. Expected: {expected_regions}, Found: {sheet2_data}")
                else:
                    # No combined.csv to cross-check — just verify at least 4 known regions present
                    known_regions = {'Central', 'East', 'North', 'South', 'West'}
                    found_regions = set(sheet2_data.keys())
                    if len(found_regions & known_regions) >= 4:
                        print(f"PASS: Component 4 — Sheet2 has region summary with {len(found_regions)} regions (0.30 pts)")
                        total_score += 0.30
                    else:
                        print(f"FAIL: Component 4 — Sheet2 region data incomplete: {found_regions}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


verify_task()
