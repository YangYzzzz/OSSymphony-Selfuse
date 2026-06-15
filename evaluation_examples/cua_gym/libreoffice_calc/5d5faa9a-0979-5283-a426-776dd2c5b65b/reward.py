"""
Reward Script: Calculate EOMONTH-based payroll next pay dates and days until pay
Task ID: calc_hr_eom_payroll_061
Domain: libreoffice_calc

Scoring Rubric (total = 1.0):
  Component 1: EOMONTH formulas in E2:E67                  — 0.35 pts
  Component 2: Days-until-pay formulas in F2:F67            — 0.25 pts
  Component 3: Date format DD/MM/YYYY on E2:E67             — 0.15 pts
  Component 4: Integer format '0' on F2:F67                 — 0.10 pts
  Component 5: Conditional formatting on F2:F67 (<=5, green)— 0.15 pts
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_eom_payroll_061'
SHEET_NAME = 'Payroll Schedule'
DATA_ROWS = range(2, 68)  # rows 2-67 inclusive (66 rows)
TOTAL_DATA_ROWS = 66


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sheet exists
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # ------------------------------------------------------------------
    # Component 1: EOMONTH formulas in E2:E67 (0.35 points)
    # Each row should have =EOMONTH(Dx,1) in column E
    # This FAILS on initial (all None) and PASSES on golden
    # ------------------------------------------------------------------
    try:
        eomonth_count = 0
        eomonth_correct = 0
        for row in DATA_ROWS:
            cell_e = ws.cell(row=row, column=5)
            val = cell_e.value
            if val is not None:
                eomonth_count += 1
                # Accept =EOMONTH(Dx,1) where x is the row number
                val_str = str(val).strip().upper().replace(' ', '')
                expected_pattern = f'=EOMONTH(D{row},1)'
                if val_str == expected_pattern.upper():
                    eomonth_correct += 1

        if eomonth_correct == TOTAL_DATA_ROWS:
            print(f"PASS: Component 1 — All {TOTAL_DATA_ROWS} EOMONTH formulas correct in E2:E67 (0.35 pts)")
            total_score += 0.35
        elif eomonth_correct >= TOTAL_DATA_ROWS * 0.9:
            partial = 0.25
            print(f"PARTIAL: Component 1 — {eomonth_correct}/{TOTAL_DATA_ROWS} EOMONTH formulas correct ({partial} pts)")
            total_score += partial
        elif eomonth_correct > 0:
            partial = 0.15
            print(f"PARTIAL: Component 1 — {eomonth_correct}/{TOTAL_DATA_ROWS} EOMONTH formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No EOMONTH formulas found in E2:E67. "
                  f"Sample E2={repr(ws.cell(row=2, column=5).value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ------------------------------------------------------------------
    # Component 2: Days-until-pay formulas in F2:F67 (0.25 points)
    # Each row should have =Ex-TODAY() in column F
    # This FAILS on initial (all None) and PASSES on golden
    # ------------------------------------------------------------------
    try:
        days_correct = 0
        for row in DATA_ROWS:
            cell_f = ws.cell(row=row, column=6)
            val = cell_f.value
            if val is not None:
                val_str = str(val).strip().upper().replace(' ', '')
                expected_pattern = f'=E{row}-TODAY()'
                if val_str == expected_pattern.upper():
                    days_correct += 1

        if days_correct == TOTAL_DATA_ROWS:
            print(f"PASS: Component 2 — All {TOTAL_DATA_ROWS} days-until-pay formulas correct in F2:F67 (0.25 pts)")
            total_score += 0.25
        elif days_correct >= TOTAL_DATA_ROWS * 0.9:
            partial = 0.18
            print(f"PARTIAL: Component 2 — {days_correct}/{TOTAL_DATA_ROWS} days formulas correct ({partial} pts)")
            total_score += partial
        elif days_correct > 0:
            partial = 0.10
            print(f"PARTIAL: Component 2 — {days_correct}/{TOTAL_DATA_ROWS} days formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No days-until-pay formulas found in F2:F67. "
                  f"Sample F2={repr(ws.cell(row=2, column=6).value)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ------------------------------------------------------------------
    # Component 3: Date format DD/MM/YYYY on E2:E67 (0.15 points)
    # Initial file has 'General' format — golden has 'DD/MM/YYYY'
    # This FAILS on initial and PASSES on golden
    # ------------------------------------------------------------------
    try:
        date_format_count = 0
        for row in DATA_ROWS:
            cell_e = ws.cell(row=row, column=5)
            fmt = cell_e.number_format
            # Accept DD/MM/YYYY (case-insensitive)
            if fmt and fmt.upper() in ('DD/MM/YYYY', 'DD/MM/YY'):
                date_format_count += 1

        if date_format_count == TOTAL_DATA_ROWS:
            print(f"PASS: Component 3 — All {TOTAL_DATA_ROWS} E-column cells have DD/MM/YYYY format (0.15 pts)")
            total_score += 0.15
        elif date_format_count >= TOTAL_DATA_ROWS * 0.9:
            partial = 0.10
            print(f"PARTIAL: Component 3 — {date_format_count}/{TOTAL_DATA_ROWS} cells have date format ({partial} pts)")
            total_score += partial
        elif date_format_count > 0:
            partial = 0.05
            print(f"PARTIAL: Component 3 — {date_format_count}/{TOTAL_DATA_ROWS} cells have date format ({partial} pts)")
            total_score += partial
        else:
            sample_fmt = ws.cell(row=2, column=5).number_format
            print(f"FAIL: Component 3 — E column not formatted as DD/MM/YYYY. Sample E2 format={repr(sample_fmt)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ------------------------------------------------------------------
    # Component 4: Integer format '0' on F2:F67 (0.10 points)
    # Initial file has 'General' format — golden has '0'
    # This FAILS on initial and PASSES on golden
    # ------------------------------------------------------------------
    try:
        int_format_count = 0
        for row in DATA_ROWS:
            cell_f = ws.cell(row=row, column=6)
            fmt = cell_f.number_format
            if fmt == '0':
                int_format_count += 1

        if int_format_count == TOTAL_DATA_ROWS:
            print(f"PASS: Component 4 — All {TOTAL_DATA_ROWS} F-column cells have integer format '0' (0.10 pts)")
            total_score += 0.10
        elif int_format_count >= TOTAL_DATA_ROWS * 0.9:
            partial = 0.07
            print(f"PARTIAL: Component 4 — {int_format_count}/{TOTAL_DATA_ROWS} cells have integer format ({partial} pts)")
            total_score += partial
        elif int_format_count > 0:
            partial = 0.04
            print(f"PARTIAL: Component 4 — {int_format_count}/{TOTAL_DATA_ROWS} cells have integer format ({partial} pts)")
            total_score += partial
        else:
            sample_fmt = ws.cell(row=2, column=6).number_format
            print(f"FAIL: Component 4 — F column not formatted as '0'. Sample F2 format={repr(sample_fmt)}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ------------------------------------------------------------------
    # Component 5: Conditional formatting on F2:F67 (0.15 points)
    # Rule: F <= 5 → background green #70AD47
    # Initial has NO conditional formatting — golden has this rule
    # This FAILS on initial and PASSES on golden
    # ------------------------------------------------------------------
    try:
        cf_found = False
        cf_correct_operator = False
        cf_correct_color = False

        cf_rules = ws.conditional_formatting
        for cf_range_obj in cf_rules:
            range_str = str(cf_range_obj)
            # Check if this CF applies to F2:F67 (or overlapping)
            if 'F' in range_str:
                for rule in cf_rules[cf_range_obj]:
                    cf_found = True
                    # Check operator is <=5 (cellIs lessThanOrEqual 5)
                    rule_type = getattr(rule, 'type', None)
                    rule_operator = getattr(rule, 'operator', None)
                    rule_formula = getattr(rule, 'formula', [])
                    if rule_type == 'cellIs' and rule_operator == 'lessThanOrEqual':
                        if rule_formula and str(rule_formula[0]).strip() == '5':
                            cf_correct_operator = True
                    # Check fill color = #70AD47 (ARGB: FF70AD47)
                    try:
                        dxf = rule.dxf
                        if dxf is not None and dxf.fill is not None:
                            fg_color = dxf.fill.fgColor.rgb
                            # Accept both FF70AD47 and 70AD47 (with or without alpha prefix)
                            if fg_color and '70AD47' in fg_color.upper():
                                cf_correct_color = True
                    except Exception:
                        pass

        if cf_found and cf_correct_operator and cf_correct_color:
            print(f"PASS: Component 5 — Conditional formatting found: F<=5 → green #70AD47 (0.15 pts)")
            total_score += 0.15
        elif cf_found and (cf_correct_operator or cf_correct_color):
            partial = 0.08
            print(f"PARTIAL: Component 5 — CF found but incomplete. "
                  f"operator_ok={cf_correct_operator}, color_ok={cf_correct_color} ({partial} pts)")
            total_score += partial
        elif cf_found:
            partial = 0.04
            print(f"PARTIAL: Component 5 — CF found on F column but rules do not match expected ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — No conditional formatting found on F column")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {round(final_score, 4)}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
