"""
Initial Setup: HR Payroll Schedule with EOMONTH task
Task ID: calc_hr_eom_payroll_061
Domain: libreoffice_calc

Creates a spreadsheet with employee payroll schedule data.
Columns E (Next Pay Date) and F (Days Until Pay) are intentionally left empty
so the agent can fill them with EOMONTH formulas.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date, timedelta
import calendar

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_eom_payroll_061'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def last_day_of_month(year, month):
    """Return the last day of the given month."""
    return calendar.monthrange(year, month)[1]


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Payroll Schedule ---
    ws = wb.active
    ws.title = 'Payroll Schedule'

    # Headers
    headers = ['Emp ID', 'Name', 'Pay Frequency', 'Last Pay Date', 'Next Pay Date', 'Days Until Pay']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Realistic employee data — 66 employees (rows 2-67)
    employees = [
        ('EMP001', 'Sarah Chen', 'Monthly'),
        ('EMP002', 'Marcus Johnson', 'Monthly'),
        ('EMP003', 'Priya Patel', 'Monthly'),
        ('EMP004', 'David Kim', 'Monthly'),
        ('EMP005', 'Olivia Martinez', 'Monthly'),
        ('EMP006', 'James Thompson', 'Monthly'),
        ('EMP007', 'Aisha Williams', 'Monthly'),
        ('EMP008', 'Nathan Rodriguez', 'Monthly'),
        ('EMP009', 'Emily Zhang', 'Monthly'),
        ('EMP010', 'Carlos Hernandez', 'Monthly'),
        ('EMP011', 'Mei Lin', 'Monthly'),
        ('EMP012', 'Brandon Lee', 'Monthly'),
        ('EMP013', 'Fatima Al-Hassan', 'Monthly'),
        ('EMP014', 'Ryan O\'Brien', 'Monthly'),
        ('EMP015', 'Isabella Nguyen', 'Monthly'),
        ('EMP016', 'Michael Brown', 'Monthly'),
        ('EMP017', 'Yuki Tanaka', 'Monthly'),
        ('EMP018', 'Destiny Jackson', 'Monthly'),
        ('EMP019', 'Hassan Ahmed', 'Monthly'),
        ('EMP020', 'Chloe Anderson', 'Monthly'),
        ('EMP021', 'Diego Reyes', 'Monthly'),
        ('EMP022', 'Samantha White', 'Monthly'),
        ('EMP023', 'Kevin Park', 'Monthly'),
        ('EMP024', 'Layla Ibrahim', 'Monthly'),
        ('EMP025', 'Tyler Scott', 'Monthly'),
        ('EMP026', 'Nadia Petrov', 'Monthly'),
        ('EMP027', 'Andre Wilson', 'Monthly'),
        ('EMP028', 'Grace Liu', 'Monthly'),
        ('EMP029', 'Patrick Murphy', 'Monthly'),
        ('EMP030', 'Sofia Gonzalez', 'Monthly'),
        ('EMP031', 'Omar Farouk', 'Monthly'),
        ('EMP032', 'Hannah Evans', 'Monthly'),
        ('EMP033', 'Kai Yamamoto', 'Monthly'),
        ('EMP034', 'Brianna Clark', 'Monthly'),
        ('EMP035', 'Luca Bianchi', 'Monthly'),
        ('EMP036', 'Tiffany Davis', 'Monthly'),
        ('EMP037', 'Ravi Sharma', 'Monthly'),
        ('EMP038', 'Alexis Turner', 'Monthly'),
        ('EMP039', 'Mohammed Al-Rashid', 'Monthly'),
        ('EMP040', 'Victoria Hall', 'Monthly'),
        ('EMP041', 'Jonathan Brooks', 'Monthly'),
        ('EMP042', 'Mia Cooper', 'Monthly'),
        ('EMP043', 'Elijah Ward', 'Monthly'),
        ('EMP044', 'Amelia Rivera', 'Monthly'),
        ('EMP045', 'Noah Phillips', 'Monthly'),
        ('EMP046', 'Zoe Campbell', 'Monthly'),
        ('EMP047', 'Ethan Mitchell', 'Monthly'),
        ('EMP048', 'Jasmine Carter', 'Monthly'),
        ('EMP049', 'Aaron Roberts', 'Monthly'),
        ('EMP050', 'Lily Thomas', 'Monthly'),
        ('EMP051', 'Sean Walker', 'Monthly'),
        ('EMP052', 'Aaliyah Young', 'Monthly'),
        ('EMP053', 'Connor King', 'Monthly'),
        ('EMP054', 'Penelope Wright', 'Monthly'),
        ('EMP055', 'Jayden Hill', 'Monthly'),
        ('EMP056', 'Aurora Scott', 'Monthly'),
        ('EMP057', 'Miles Green', 'Monthly'),
        ('EMP058', 'Stella Adams', 'Monthly'),
        ('EMP059', 'Dominic Baker', 'Monthly'),
        ('EMP060', 'Naomi Nelson', 'Monthly'),
        ('EMP061', 'Caleb Carter', 'Monthly'),
        ('EMP062', 'Leah Mitchell', 'Monthly'),
        ('EMP063', 'Isaiah Perez', 'Monthly'),
        ('EMP064', 'Scarlett Roberts', 'Monthly'),
        ('EMP065', 'Zachary Turner', 'Monthly'),
        ('EMP066', 'Audrey Phillips', 'Monthly'),
    ]

    # Vary last pay dates across different months (all end-of-month dates)
    # Use a mix of months ending in 2024/2025 — all are "last pay date" (end of previous month)
    last_pay_dates = []
    base_dates = [
        date(2025, 1, 31),
        date(2025, 1, 31),
        date(2024, 12, 31),
        date(2025, 1, 31),
        date(2024, 11, 30),
        date(2025, 1, 31),
        date(2024, 12, 31),
        date(2025, 1, 31),
        date(2024, 10, 31),
        date(2025, 1, 31),
        date(2024, 12, 31),
        date(2025, 1, 31),
        date(2024, 11, 30),
        date(2025, 1, 31),
        date(2024, 12, 31),
        date(2025, 1, 31),
        date(2024, 12, 31),
        date(2025, 1, 31),
        date(2024, 11, 30),
        date(2025, 1, 31),
        date(2024, 12, 31),
        date(2025, 1, 31),
        date(2025, 1, 31),
        date(2024, 12, 31),
        date(2025, 1, 31),
        date(2024, 11, 30),
        date(2025, 1, 31),
        date(2024, 12, 31),
        date(2025, 1, 31),
        date(2024, 12, 31),
        date(2025, 1, 31),
        date(2024, 11, 30),
        date(2024, 12, 31),
        date(2025, 1, 31),
        date(2024, 12, 31),
        date(2025, 1, 31),
        date(2024, 12, 31),
        date(2025, 1, 31),
        date(2024, 11, 30),
        date(2025, 1, 31),
        date(2024, 12, 31),
        date(2025, 1, 31),
        date(2025, 1, 31),
        date(2024, 12, 31),
        date(2025, 1, 31),
        date(2024, 12, 31),
        date(2025, 1, 31),
        date(2024, 11, 30),
        date(2025, 1, 31),
        date(2024, 12, 31),
        date(2025, 1, 31),
        date(2025, 1, 31),
        date(2024, 12, 31),
        date(2025, 1, 31),
        date(2024, 12, 31),
        date(2025, 1, 31),
        date(2024, 11, 30),
        date(2025, 1, 31),
        date(2024, 12, 31),
        date(2025, 1, 31),
        date(2024, 12, 31),
        date(2025, 1, 31),
        date(2024, 11, 30),
        date(2025, 1, 31),
        date(2024, 12, 31),
        date(2025, 1, 31),
    ]

    for i, (emp_id, name, freq) in enumerate(employees):
        row = i + 2
        last_pay = base_dates[i]

        ws.cell(row=row, column=1, value=emp_id)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=freq)

        # Set last pay date as a date value with date format
        date_cell = ws.cell(row=row, column=4, value=last_pay)
        date_cell.number_format = 'DD/MM/YYYY'

        # Columns E (Next Pay Date) and F (Days Until Pay) MUST be empty
        ws.cell(row=row, column=5, value=None)
        ws.cell(row=row, column=6, value=None)

    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Payroll Schedule')
    print(f'  Rows: 67 (1 header + 66 employees)')
    print(f'  Columns E & F: empty (to be filled by agent)')


create_initial()
