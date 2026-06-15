"""
Initial Setup: Utility Bills Tracker with PDF bills on desktop
Task ID: osworld_multi_apps_receipt_to_calc_003
Domain: libreoffice_calc

Creates:
  - /home/user/osworld_multi_apps_receipt_to_calc_003.xlsx (bills_tracker.xlsx)
  - /home/user/Desktop/electric_bill.pdf
  - /home/user/Desktop/water_bill.pdf
  - /home/user/Desktop/internet_bill.pdf
Opens LibreOffice Calc with the xlsx file.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_receipt_to_calc_003'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'
DESKTOP = f'{WORKDIR}/Desktop'

# ---- Bill data used in PDFs (must match golden patch) ----
ELECTRIC_AMOUNT = 127.45
ELECTRIC_DUE    = '2026-03-25'
WATER_AMOUNT    = 43.80
WATER_DUE       = '2026-03-28'
INTERNET_AMOUNT = 89.99
INTERNET_DUE    = '2026-03-22'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_bills_tracker():
    """Create bills_tracker.xlsx with existing previous month data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bills"

    # --- Column widths ---
    ws.column_dimensions['A'].width = 12  # Month
    ws.column_dimensions['B'].width = 16  # Utility
    ws.column_dimensions['C'].width = 12  # Amount
    ws.column_dimensions['D'].width = 14  # Due Date
    ws.column_dimensions['E'].width = 8   # Paid

    # --- Header row ---
    headers = ['Month', 'Utility', 'Amount', 'Due Date', 'Paid']
    header_font = Font(name='Calibri', bold=True, size=11)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(name='Calibri', bold=True, size=11, color='FFFFFFFF')
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border

    ws.row_dimensions[1].height = 20

    # --- Previous months data (Jan and Feb 2026) ---
    previous_data = [
        # Month,    Utility,   Amount,  Due Date,      Paid
        ['January', 'Electric', 118.30, '2026-01-22',  'Yes'],
        ['January', 'Water',    39.50,  '2026-01-28',  'Yes'],
        ['January', 'Internet', 89.99,  '2026-01-20',  'Yes'],
        ['February','Electric', 134.75, '2026-02-24',  'Yes'],
        ['February','Water',    42.10,  '2026-02-26',  'Yes'],
        ['February','Internet', 89.99,  '2026-02-21',  'Yes'],
    ]

    data_alignment = Alignment(horizontal='left', vertical='center')
    for row_idx, row_data in enumerate(previous_data, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name='Calibri', size=11)
            cell.alignment = data_alignment
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            # Format Amount as currency
            if col_idx == 3:
                cell.number_format = '$#,##0.00'

    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


def create_pdf_bill(filepath, title, company, account_number, service_period,
                    amount, due_date, address):
    """Create a simple PDF utility bill using fpdf2."""
    try:
        from fpdf import FPDF

        pdf = FPDF()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        # Header
        pdf.set_font('Helvetica', 'B', 20)
        pdf.set_fill_color(44, 86, 148)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(0, 14, title, new_x='LMARGIN', new_y='NEXT', align='C', fill=True)
        pdf.ln(4)

        # Customer info
        pdf.set_text_color(0, 0, 0)
        pdf.set_font('Helvetica', '', 11)
        pdf.cell(0, 7, f'Company: {company}', new_x='LMARGIN', new_y='NEXT')
        pdf.cell(0, 7, f'Account Number: {account_number}', new_x='LMARGIN', new_y='NEXT')
        pdf.cell(0, 7, f'Service Address: {address}', new_x='LMARGIN', new_y='NEXT')
        pdf.cell(0, 7, f'Billing Period: {service_period}', new_x='LMARGIN', new_y='NEXT')
        pdf.ln(6)

        # Bill summary box
        pdf.set_fill_color(240, 240, 240)
        pdf.set_font('Helvetica', 'B', 13)
        pdf.cell(0, 10, 'BILLING SUMMARY', new_x='LMARGIN', new_y='NEXT', fill=True)
        pdf.ln(2)

        pdf.set_font('Helvetica', '', 12)
        pdf.cell(90, 8, 'Total Amount Due:', new_x='RIGHT', new_y='TOP')
        pdf.set_font('Helvetica', 'B', 12)
        pdf.cell(0, 8, f'${amount:.2f}', new_x='LMARGIN', new_y='NEXT')
        pdf.ln(2)

        pdf.set_font('Helvetica', '', 12)
        pdf.cell(90, 8, 'Payment Due Date:', new_x='RIGHT', new_y='TOP')
        pdf.set_font('Helvetica', 'B', 12)
        pdf.cell(0, 8, due_date, new_x='LMARGIN', new_y='NEXT')
        pdf.ln(8)

        pdf.set_font('Helvetica', 'I', 10)
        pdf.set_text_color(100, 100, 100)
        pdf.multi_cell(0, 6,
            'Please pay by the due date to avoid late fees. '
            'Questions? Call 1-800-555-0100 or visit our website.')

        pdf.output(filepath)
        print(f'PDF created: {filepath}')
    except Exception as e:
        # Fallback: create a minimal text-based PDF using plain bytes if fpdf2 unavailable
        print(f'fpdf2 error ({e}), writing minimal PDF fallback.')
        content = (
            f'%PDF-1.4\n'
            f'1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n'
            f'2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n'
            f'3 0 obj<</Type/Page/MediaBox[0 0 612 792]/Parent 2 0 R/Resources<</Font<</F1 4 0 R>>>>>>\nendobj\n'
            f'4 0 obj<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>endobj\n'
            f'5 0 obj<</Length 200>>\nstream\n'
            f'BT /F1 14 Tf 50 750 Td ({title}) Tj\n'
            f'/F1 12 Tf 0 -30 Td (Amount Due: ${amount:.2f}) Tj\n'
            f'/F1 12 Tf 0 -20 Td (Due Date: {due_date}) Tj\n'
            f'ET\nendstream\nendobj\n'
            f'xref\n0 6\n0000000000 65535 f \n'
            f'trailer<</Size 6/Root 1 0 R>>\n%%EOF\n'
        )
        with open(filepath, 'w') as f:
            f.write(content)
        print(f'Minimal PDF fallback written: {filepath}')


def create_pdf_bills():
    """Create the three utility bill PDFs on the Desktop."""
    os.makedirs(DESKTOP, exist_ok=True)

    create_pdf_bill(
        filepath=f'{DESKTOP}/electric_bill.pdf',
        title='ELECTRIC UTILITY BILL',
        company='CityLight Electric Co.',
        account_number='EL-20240892',
        service_period='March 1 - March 31, 2026',
        amount=ELECTRIC_AMOUNT,
        due_date=ELECTRIC_DUE,
        address='742 Evergreen Terrace, Springfield',
    )

    create_pdf_bill(
        filepath=f'{DESKTOP}/water_bill.pdf',
        title='WATER & SEWER BILL',
        company='Metro Water Services',
        account_number='WS-00571143',
        service_period='March 1 - March 31, 2026',
        amount=WATER_AMOUNT,
        due_date=WATER_DUE,
        address='742 Evergreen Terrace, Springfield',
    )

    create_pdf_bill(
        filepath=f'{DESKTOP}/internet_bill.pdf',
        title='INTERNET SERVICE BILL',
        company='FastConnect ISP',
        account_number='ISP-88241-B',
        service_period='March 1 - March 31, 2026',
        amount=INTERNET_AMOUNT,
        due_date=INTERNET_DUE,
        address='742 Evergreen Terrace, Springfield',
    )


def main():
    create_bills_tracker()
    create_pdf_bills()

    # GUI-ready startup: open LibreOffice Calc with the bills tracker
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


main()
