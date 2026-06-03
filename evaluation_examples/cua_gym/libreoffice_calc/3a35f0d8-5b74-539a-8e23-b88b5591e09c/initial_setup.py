"""
Initial Setup: Add headers/footers to audit report spreadsheet
Task ID: calc_gfl_050
Domain: libreoffice_calc

Creates a Transactions sheet with 45 rows of audit data.
No headers or footers are configured (that is the agent's task).
Opens the file in LibreOffice Calc for GUI-ready state.
"""

import os
import shlex
import subprocess
import time
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from datetime import datetime, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_050'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Transactions'

    # --- Headers ---
    headers = ['Transaction ID', 'Date', 'Account', 'Debit', 'Credit', 'Balance', 'Verified']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Column widths ---
    ws.column_dimensions['A'].width = 18  # Transaction ID
    ws.column_dimensions['B'].width = 14  # Date
    ws.column_dimensions['C'].width = 28  # Account
    ws.column_dimensions['D'].width = 14  # Debit
    ws.column_dimensions['E'].width = 14  # Credit
    ws.column_dimensions['F'].width = 14  # Balance
    ws.column_dimensions['G'].width = 12  # Verified

    # --- Realistic transaction data (45 rows) ---
    accounts = [
        'Office Supplies', 'Payroll - Engineering', 'Marketing Expense',
        'Travel & Entertainment', 'Software Licenses', 'Consulting Fees',
        'Utilities', 'Rent - Main Office', 'Insurance Premium',
        'Equipment Maintenance', 'Client Reimbursement', 'Training & Dev',
        'Legal Services', 'Telecommunications', 'Payroll - Marketing',
        'Cloud Infrastructure', 'Office Furniture', 'Employee Benefits',
        'Professional Services', 'Petty Cash', 'Accounts Receivable',
        'Revenue - Product Sales', 'Revenue - Services', 'Payroll - Sales',
        'Shipping & Logistics',
    ]

    random.seed(42)
    base_date = datetime(2025, 1, 6)
    balance = 125000.00
    data_font = Font(name='Calibri', size=11)
    date_format = 'yyyy-mm-dd'
    money_format = '#,##0.00'

    for i in range(1, 46):
        row = i + 1
        # Transaction ID
        tid = f'TXN-2025-{i:04d}'
        ws.cell(row=row, column=1, value=tid).font = data_font

        # Date (spread across Jan-Mar 2025)
        tx_date = base_date + timedelta(days=random.randint(0, 80))
        cell_date = ws.cell(row=row, column=2, value=tx_date)
        cell_date.number_format = date_format
        cell_date.font = data_font

        # Account
        acct = random.choice(accounts)
        ws.cell(row=row, column=3, value=acct).font = data_font

        # Debit / Credit (one is filled, other is 0 or blank)
        is_debit = random.random() < 0.6
        if is_debit:
            debit_val = round(random.uniform(150, 18500), 2)
            credit_val = 0.0
            balance -= debit_val
        else:
            debit_val = 0.0
            credit_val = round(random.uniform(500, 25000), 2)
            balance += credit_val

        cell_d = ws.cell(row=row, column=4, value=debit_val if debit_val > 0 else None)
        cell_d.number_format = money_format
        cell_d.font = data_font

        cell_c = ws.cell(row=row, column=5, value=credit_val if credit_val > 0 else None)
        cell_c.number_format = money_format
        cell_c.font = data_font

        cell_b = ws.cell(row=row, column=6, value=round(balance, 2))
        cell_b.number_format = money_format
        cell_b.font = data_font

        # Verified
        verified = random.choice(['Yes', 'Yes', 'Yes', 'Pending', 'No'])
        ws.cell(row=row, column=7, value=verified).font = data_font

    # Freeze header row
    ws.freeze_panes = 'A2'

    # --- Explicitly ensure NO headers or footers ---
    ws.oddHeader.left.text = ''
    ws.oddHeader.center.text = ''
    ws.oddHeader.right.text = ''
    ws.oddFooter.left.text = ''
    ws.oddFooter.center.text = ''
    ws.oddFooter.right.text = ''

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
