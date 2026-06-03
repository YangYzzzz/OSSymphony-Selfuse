"""
Initial Setup: Pivot table showing COUNT of sales, needs to be changed to SUM of Revenue
Task ID: calc_pivot_021
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_021'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

# --- Data generation ---
PRODUCTS = [
    'Laptop Pro 15', 'Wireless Mouse', 'USB-C Hub', 'Mechanical Keyboard',
    'Monitor 27in', '4TB External SSD', 'Webcam HD', 'Desk Lamp LED',
    'Noise-Cancel Headphones', 'Tablet Stand'
]
REGIONS = ['North', 'South', 'East', 'West', 'Central']

def launch_gui(command: str, delay_sec: float = 1.0):
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()

    # ========== Sheet 1: Data ==========
    ws_data = wb.active
    ws_data.title = 'Data'

    headers = ['ID', 'Product', 'Region', 'Revenue']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")

    for col, h in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align

    # Generate 150 rows with total revenue = 225000
    # Strategy: generate raw values then scale to hit exact target
    rows_data = []
    target_total = 225000

    # First pass: generate products, regions, and raw revenue
    raw_rows = []
    for i in range(1, 151):
        product = random.choice(PRODUCTS)
        region = random.choice(REGIONS)
        raw_revenue = random.uniform(500, 3000)
        raw_rows.append([i, product, region, raw_revenue])

    # Scale revenues to hit exact target
    raw_total = sum(r[3] for r in raw_rows)
    scale = target_total / raw_total
    for r in raw_rows:
        r[3] = round(r[3] * scale, 2)

    # Adjust last row to fix rounding difference
    current_total = sum(r[3] for r in raw_rows)
    raw_rows[-1][3] = round(raw_rows[-1][3] + (target_total - current_total), 2)
    rows_data = raw_rows

    for r, row_data in enumerate(rows_data, 2):
        ws_data.cell(row=r, column=1, value=row_data[0])   # ID
        ws_data.cell(row=r, column=2, value=row_data[1])    # Product
        ws_data.cell(row=r, column=3, value=row_data[2])    # Region
        ws_data.cell(row=r, column=4, value=row_data[3])    # Revenue

    # Column widths
    ws_data.column_dimensions['A'].width = 8
    ws_data.column_dimensions['B'].width = 28
    ws_data.column_dimensions['C'].width = 12
    ws_data.column_dimensions['D'].width = 14

    # Format revenue column as currency
    for r in range(2, 152):
        ws_data.cell(row=r, column=4).number_format = '$#,##0.00'

    # ========== Sheet 2: PivotResult — COUNT of ID ==========
    ws_pivot = wb.create_sheet('PivotResult')

    # Build cross-tabulation: COUNT of ID by Product x Region
    from collections import defaultdict
    count_table = defaultdict(lambda: defaultdict(int))

    for row_data in rows_data:
        product = row_data[1]
        region = row_data[2]
        count_table[product][region] += 1

    sorted_products = sorted(count_table.keys())

    # Title
    ws_pivot.cell(row=1, column=1, value='COUNT of ID')
    ws_pivot.cell(row=1, column=1).font = Font(bold=True, size=12)

    # Header row: blank + regions + Grand Total
    pivot_header_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
    ws_pivot.cell(row=2, column=1, value='Product')
    ws_pivot.cell(row=2, column=1).font = Font(bold=True)
    ws_pivot.cell(row=2, column=1).fill = pivot_header_fill
    for ci, reg in enumerate(REGIONS, 2):
        cell = ws_pivot.cell(row=2, column=ci, value=reg)
        cell.font = Font(bold=True)
        cell.fill = pivot_header_fill
        cell.alignment = Alignment(horizontal="center")
    gt_col = len(REGIONS) + 2
    cell = ws_pivot.cell(row=2, column=gt_col, value='Grand Total')
    cell.font = Font(bold=True)
    cell.fill = pivot_header_fill
    cell.alignment = Alignment(horizontal="center")

    # Data rows
    for ri, product in enumerate(sorted_products, 3):
        ws_pivot.cell(row=ri, column=1, value=product)
        ws_pivot.cell(row=ri, column=1).font = Font(bold=False)
        row_total = 0
        for ci, reg in enumerate(REGIONS, 2):
            val = count_table[product][reg]
            ws_pivot.cell(row=ri, column=ci, value=val)
            ws_pivot.cell(row=ri, column=ci).alignment = Alignment(horizontal="center")
            row_total += val
        ws_pivot.cell(row=ri, column=gt_col, value=row_total)
        ws_pivot.cell(row=ri, column=gt_col).alignment = Alignment(horizontal="center")

    # Grand Total row
    total_row = 3 + len(sorted_products)
    ws_pivot.cell(row=total_row, column=1, value='Grand Total')
    ws_pivot.cell(row=total_row, column=1).font = Font(bold=True)
    ws_pivot.cell(row=total_row, column=1).fill = pivot_header_fill
    for ci, reg in enumerate(REGIONS, 2):
        col_total = sum(count_table[p][reg] for p in sorted_products)
        cell = ws_pivot.cell(row=total_row, column=ci, value=col_total)
        cell.font = Font(bold=True)
        cell.fill = pivot_header_fill
        cell.alignment = Alignment(horizontal="center")
    grand_total = 150
    cell = ws_pivot.cell(row=total_row, column=gt_col, value=grand_total)
    cell.font = Font(bold=True)
    cell.fill = pivot_header_fill
    cell.alignment = Alignment(horizontal="center")

    # Column widths for pivot
    ws_pivot.column_dimensions['A'].width = 28
    for ci in range(2, gt_col + 1):
        col_letter = openpyxl.utils.get_column_letter(ci)
        ws_pivot.column_dimensions[col_letter].width = 14

    # Thin borders for the pivot table area
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(2, total_row + 1):
        for c in range(1, gt_col + 1):
            ws_pivot.cell(row=r, column=c).border = border

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
