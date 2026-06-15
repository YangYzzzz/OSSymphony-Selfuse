"""
Reward Script: Change pivot table aggregation from COUNT of ID to SUM of Revenue
Task ID: calc_pivot_021
Domain: libreoffice_calc
Scoring:
  Component 1: A1 header changed to "SUM of Revenue" (0.25 pts)
  Component 2: Data cells contain revenue values, not counts (0.35 pts)
  Component 3: Grand Total equals 225000 (0.25 pts)
  Component 4: Layout preserved — same products as rows, regions as columns (0.15 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_021'

# Expected layout from task context
EXPECTED_PRODUCTS = [
    '4TB External SSD', 'Desk Lamp LED', 'Laptop Pro 15',
    'Mechanical Keyboard', 'Monitor 27in', 'Noise-Cancel Headphones',
    'Tablet Stand', 'USB-C Hub', 'Webcam HD', 'Wireless Mouse'
]
EXPECTED_REGIONS = ['North', 'South', 'East', 'West', 'Central']
EXPECTED_GRAND_TOTAL = 225000.0


def persist_app_state(domain: str):
    """Save any unsaved LibreOffice state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check PivotResult sheet exists (precondition gate)
    if 'PivotResult' not in wb.sheetnames:
        print("CRITICAL: 'PivotResult' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['PivotResult']

    # Component 1: A1 header changed to "SUM of Revenue" (0.25 points)
    # Initial has "COUNT of ID", golden should have "SUM of Revenue"
    try:
        a1_value = ws['A1'].value
        if a1_value and 'SUM' in str(a1_value).upper() and 'REVENUE' in str(a1_value).upper():
            print(f"PASS: Component 1 — A1 header is '{a1_value}' (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — Expected 'SUM of Revenue' in A1, found: '{a1_value}'")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Data cells contain revenue values (floats), not integer counts (0.35 points)
    # In initial, data values are small integers (counts). In golden, they should be float revenue values.
    # We check a sample of data cells (B3:F5) — they should be floats > 100 (revenue) not small ints (counts)
    try:
        revenue_cell_count = 0
        total_data_cells = 0
        for row_idx in range(3, 13):  # rows 3-12 are product rows
            for col_idx in range(2, 7):  # columns B-F are region columns
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                total_data_cells += 1
                if cell_val is not None and isinstance(cell_val, (int, float)):
                    # Revenue values should generally be > 100 or == 0
                    # Count values are small integers (0-7)
                    # A cell with value > 50 is almost certainly a revenue value, not a count
                    if float(cell_val) > 50 or float(cell_val) == 0:
                        revenue_cell_count += 1

        # Need most data cells to have revenue-scale values
        # In initial, most counts are 0-7, so very few would pass > 50
        # In golden, revenue values are typically 500-13000+
        ratio = revenue_cell_count / total_data_cells if total_data_cells > 0 else 0
        if ratio >= 0.8:
            print(f"PASS: Component 2 — {revenue_cell_count}/{total_data_cells} data cells have revenue-scale values (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 2 — Only {revenue_cell_count}/{total_data_cells} data cells have revenue-scale values (ratio={ratio:.2f})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Grand Total equals 225000 (0.25 points)
    # Initial Grand Total is 150 (count), golden should be 225000
    try:
        grand_total = ws['G13'].value
        if grand_total is not None:
            gt_float = float(grand_total)
            if abs(gt_float - EXPECTED_GRAND_TOTAL) < 1.0:
                print(f"PASS: Component 3 — Grand Total G13 = {gt_float} (expected ~{EXPECTED_GRAND_TOTAL}) (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 3 — Grand Total G13 = {gt_float}, expected ~{EXPECTED_GRAND_TOTAL}")
        else:
            print(f"FAIL: Component 3 — Grand Total G13 is None")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Layout preserved — same products as rows, same regions as columns (0.15 points)
    # Both initial and golden should have same layout, but this component only awards points
    # when combined with the aggregation change (A1 must NOT be "COUNT of ID")
    try:
        # Check that A1 is NOT "COUNT of ID" (i.e., the change happened) AND layout is preserved
        a1_val = str(ws['A1'].value) if ws['A1'].value else ''
        if 'COUNT' in a1_val.upper():
            # No change happened — this is still initial state
            print(f"FAIL: Component 4 — Aggregation not changed (A1 still says '{a1_val}'), layout check skipped")
        else:
            # Check product names in column A (rows 3-12)
            products_found = []
            for r in range(3, 13):
                val = ws.cell(row=r, column=1).value
                if val:
                    products_found.append(str(val).strip())

            # Check region headers in row 2 (columns B-F)
            regions_found = []
            for c in range(2, 7):
                val = ws.cell(row=2, column=c).value
                if val:
                    regions_found.append(str(val).strip())

            products_match = sorted(products_found) == sorted(EXPECTED_PRODUCTS)
            regions_match = sorted(regions_found) == sorted(EXPECTED_REGIONS)

            if products_match and regions_match:
                print(f"PASS: Component 4 — Layout preserved: {len(products_found)} products, {len(regions_found)} regions (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 4 — Layout mismatch. Products match: {products_match}, Regions match: {regions_match}")
                if not products_match:
                    print(f"  Expected products: {sorted(EXPECTED_PRODUCTS)}")
                    print(f"  Found products: {sorted(products_found)}")
                if not regions_match:
                    print(f"  Expected regions: {sorted(EXPECTED_REGIONS)}")
                    print(f"  Found regions: {sorted(regions_found)}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
