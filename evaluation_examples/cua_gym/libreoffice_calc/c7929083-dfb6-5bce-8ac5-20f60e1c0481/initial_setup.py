"""
Initial Setup: Portfolio tracker with stock data, no named ranges, F2 empty.
Task ID: calc_nrv_030
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_030'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Headers
    ws["A1"] = "Ticker"
    ws["B1"] = "Company"
    ws["C1"] = "Price"
    ws["D1"] = "Shares"
    ws["F1"] = "Portfolio Value"

    # 24 stocks with realistic data (rows 2-25)
    stocks = [
        ("AAPL", "Apple Inc.", 189.50, 150),
        ("MSFT", "Microsoft Corp.", 378.25, 85),
        ("GOOGL", "Alphabet Inc.", 141.80, 60),
        ("AMZN", "Amazon.com Inc.", 178.35, 45),
        ("NVDA", "NVIDIA Corp.", 495.20, 30),
        ("META", "Meta Platforms Inc.", 356.70, 40),
        ("TSLA", "Tesla Inc.", 248.90, 55),
        ("BRK.B", "Berkshire Hathaway", 362.15, 20),
        ("JPM", "JPMorgan Chase & Co.", 172.40, 75),
        ("V", "Visa Inc.", 258.30, 65),
        ("JNJ", "Johnson & Johnson", 156.85, 90),
        ("WMT", "Walmart Inc.", 165.20, 80),
        ("PG", "Procter & Gamble Co.", 148.95, 70),
        ("MA", "Mastercard Inc.", 412.60, 25),
        ("HD", "Home Depot Inc.", 345.70, 35),
        ("CVX", "Chevron Corp.", 152.40, 50),
        ("MRK", "Merck & Co. Inc.", 108.25, 110),
        ("ABBV", "AbbVie Inc.", 162.30, 95),
        ("PEP", "PepsiCo Inc.", 171.45, 60),
        ("KO", "Coca-Cola Co.", 59.80, 200),
        ("COST", "Costco Wholesale", 572.90, 15),
        ("AVGO", "Broadcom Inc.", 620.35, 18),
        ("LLY", "Eli Lilly & Co.", 582.10, 22),
        ("TMO", "Thermo Fisher Scientific", 528.45, 12),
    ]

    for r, (ticker, company, price, shares) in enumerate(stocks, 2):
        ws.cell(row=r, column=1, value=ticker)
        ws.cell(row=r, column=2, value=company)
        ws.cell(row=r, column=3, value=price)
        ws.cell(row=r, column=4, value=shares)

    # F2 is intentionally left empty (task requires agent to fill it)
    # No named ranges defined (task requires agent to create them)

    # Set reasonable column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["F"].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
