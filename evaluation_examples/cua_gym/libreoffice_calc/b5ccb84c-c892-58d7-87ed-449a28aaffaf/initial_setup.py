"""
Initial Setup: Vendor payment schedule spreadsheet
Task ID: calc_fin_payment_schedule_029
Domain: libreoffice_calc
"""

import openpyxl
from datetime import date, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_payment_schedule_029'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'PaymentSchedule'

    # Headers: Vendor (A), Invoice# (B), Due Date (C), Amount (D), Status (E)
    headers = ['Vendor', 'Invoice#', 'Due Date', 'Amount', 'Status']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Base date: today
    today = date(2026, 3, 3)

    # 39 rows of realistic vendor payment data (rows 2-40)
    # Due dates are deliberately NOT sorted - spread across past and future
    data = [
        ('Apex Office Supplies',    'INV-2026-0331', today + timedelta(days=45),  1250.00,  'Pending'),
        ('GlobalTech Solutions',    'INV-2026-0102', today + timedelta(days=3),   8750.50,  'Pending'),
        ('Meridian Logistics',      'INV-2026-0415', today - timedelta(days=10),  3420.75,  'Paid'),
        ('BlueStar Consulting',     'INV-2026-0509', today + timedelta(days=21),  5600.00,  'Pending'),
        ('Pinnacle Hardware',       'INV-2026-0617', today + timedelta(days=60),  980.25,   'Pending'),
        ('ClearPath Analytics',     'INV-2026-0728', today + timedelta(days=12),  12300.00, 'Pending'),
        ('Redwood Facilities',      'INV-2026-0833', today - timedelta(days=5),   2175.60,  'Paid'),
        ('Ironclad Security',       'INV-2026-0941', today + timedelta(days=6),   4500.00,  'Pending'),
        ('NovaBridge IT',           'INV-2026-1050', today + timedelta(days=30),  7890.00,  'Pending'),
        ('Harbor View Catering',    'INV-2026-1163', today + timedelta(days=9),   1850.00,  'Pending'),
        ('Summit Cloud Services',   'INV-2026-1271', today - timedelta(days=15),  6450.00,  'Paid'),
        ('Falcon Print Media',      'INV-2026-1382', today + timedelta(days=5),   3100.00,  'Pending'),
        ('Cascade Energy Group',    'INV-2026-1490', today + timedelta(days=18),  9200.00,  'Pending'),
        ('Delta Supply Chain',      'INV-2026-1504', today + timedelta(days=55),  1600.50,  'Pending'),
        ('Evergreen Telecom',       'INV-2026-1615', today + timedelta(days=2),   5350.00,  'Pending'),
        ('Maplewood Insurance',     'INV-2026-1723', today - timedelta(days=20),  11000.00, 'Paid'),
        ('Pacific Rim Imports',     'INV-2026-1831', today + timedelta(days=27),  4200.00,  'Pending'),
        ('Sterling Data Center',    'INV-2026-1942', today + timedelta(days=14),  8100.00,  'Pending'),
        ('Quantum Research Labs',   'INV-2026-2053', today + timedelta(days=40),  6700.00,  'Pending'),
        ('Vanguard Fleet Mgmt',     'INV-2026-2164', today + timedelta(days=7),   3750.00,  'Pending'),
        ('Cornerstone Finance',     'INV-2026-2275', today - timedelta(days=8),   2900.00,  'Paid'),
        ('Phoenix Web Design',      'INV-2026-2383', today + timedelta(days=35),  1400.00,  'Pending'),
        ('Titan Construction',      'INV-2026-2491', today + timedelta(days=11),  15500.00, 'Pending'),
        ('Aurora Healthcare',       'INV-2026-2502', today + timedelta(days=50),  4800.00,  'Pending'),
        ('Horizon Travel Agency',   'INV-2026-2613', today + timedelta(days=4),   2650.00,  'Pending'),
        ('Nexus Software Inc.',     'INV-2026-2721', today - timedelta(days=3),   7200.00,  'Paid'),
        ('Coastal Waste Mgmt',      'INV-2026-2832', today + timedelta(days=16),  1100.00,  'Pending'),
        ('Granite Legal Services',  'INV-2026-2943', today + timedelta(days=25),  5900.00,  'Pending'),
        ('Ember Marketing Group',   'INV-2026-3051', today + timedelta(days=58),  3300.00,  'Pending'),
        ('Tidal Wave Recruitment',  'INV-2026-3162', today + timedelta(days=8),   2200.00,  'Pending'),
        ('Ironwood Furniture Co.',  'INV-2026-3273', today - timedelta(days=12),  4400.00,  'Paid'),
        ('SkyBridge Architects',    'INV-2026-3381', today + timedelta(days=22),  6800.00,  'Pending'),
        ('Crystal Clear Windows',   'INV-2026-3492', today + timedelta(days=43),  880.00,   'Pending'),
        ('Northgate Pharmaceuticals','INV-2026-3503', today + timedelta(days=13), 9500.00,  'Pending'),
        ('Lakeview Hospitality',    'INV-2026-3614', today + timedelta(days=1),   3200.00,  'Pending'),
        ('BluePath Engineering',    'INV-2026-3722', today - timedelta(days=18),  5100.00,  'Paid'),
        ('Redstone Mining Corp',    'INV-2026-3833', today + timedelta(days=33),  7400.00,  'Pending'),
        ('Olympus Event Planning',  'INV-2026-3941', today + timedelta(days=20),  1950.00,  'Pending'),
        ('Starfield Innovations',   'INV-2026-4052', today + timedelta(days=48),  6100.00,  'Pending'),
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])  # Vendor
        ws.cell(row=r, column=2, value=row_data[1])  # Invoice#
        ws.cell(row=r, column=3, value=row_data[2])  # Due Date
        ws.cell(row=r, column=3).number_format = 'yyyy-mm-dd'
        ws.cell(row=r, column=4, value=row_data[3])  # Amount (plain number, no currency format)
        ws.cell(row=r, column=5, value=row_data[4])  # Status

    # Columns F and G are intentionally empty (no data, no headers)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

create_initial()
