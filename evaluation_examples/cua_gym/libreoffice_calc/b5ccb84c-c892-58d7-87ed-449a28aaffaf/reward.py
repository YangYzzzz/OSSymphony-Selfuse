"""
Reward Script: Vendor Payment Schedule Task
Task ID: calc_fin_payment_schedule_029
Domain: libreoffice_calc
Scoring:
  1. F1='Days Until Due' and F2:F40 IF formula  (0.25 pts)
  2. G1='30-Day Total' and G2 SUMIFS formula     (0.20 pts)
  3. G2 currency formatted and bold              (0.10 pts)
  4. D2:D40 currency formatted                   (0.10 pts)
  5. Table sorted by C column ascending          (0.15 pts)
  6. Conditional formatting on C2:C40            (0.20 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_payment_schedule_029'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get the PaymentSchedule sheet
    try:
        if 'PaymentSchedule' in wb.sheetnames:
            ws = wb['PaymentSchedule']
        else:
            ws = wb.worksheets[0]
    except Exception as e:
        print(f"CRITICAL: Cannot access sheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: F1='Days Until Due' header and F2:F40 IF formulas (0.25 points)
    try:
        f1_val = ws.cell(row=1, column=6).value
        f1_ok = isinstance(f1_val, str) and f1_val.strip() == 'Days Until Due'

        # Count F2:F40 cells that have an IF formula referencing TODAY()
        if_formula_count = 0
        for row in range(2, 41):
            val = ws.cell(row=row, column=6).value
            if isinstance(val, str):
                normalized = val.upper().replace(' ', '')
                if '=IF(' in normalized and 'TODAY()' in normalized:
                    if_formula_count += 1

        formulas_ok = if_formula_count >= 35  # at least 35 out of 39

        if f1_ok and formulas_ok:
            print(f"PASS: Component 1 — F1='Days Until Due' and F2:F40 IF/TODAY formulas ({if_formula_count}/39 with formula) (0.25 pts)")
            total_score += 0.25
        elif f1_ok and if_formula_count > 0:
            print(f"PARTIAL: Component 1 — F1 OK but only {if_formula_count}/39 IF formulas found (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 — F1={repr(f1_val)}, IF formulas count={if_formula_count}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: G1='30-Day Total' and G2 SUMIFS formula (0.20 points)
    try:
        g1_val = ws.cell(row=1, column=7).value
        g1_ok = isinstance(g1_val, str) and g1_val.strip() == '30-Day Total'

        g2_val = ws.cell(row=2, column=7).value
        g2_formula_ok = False
        if isinstance(g2_val, str):
            g2_norm = g2_val.upper().replace(' ', '')
            # Check for SUMIFS or SUMIF with Pending and TODAY criteria
            if 'SUMIF' in g2_norm and 'PENDING' in g2_norm and 'TODAY' in g2_norm:
                g2_formula_ok = True

        if g1_ok and g2_formula_ok:
            print(f"PASS: Component 2 — G1='30-Day Total' and G2 SUMIFS formula (0.20 pts)")
            total_score += 0.20
        elif g1_ok:
            print(f"PARTIAL: Component 2 — G1 OK but G2 formula wrong: {repr(g2_val)} (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 2 — G1={repr(g1_val)}, G2={repr(g2_val)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: G2 currency formatted and bold (0.10 points)
    try:
        g2_cell = ws.cell(row=2, column=7)
        g2_bold = g2_cell.font.bold is True
        g2_format = g2_cell.number_format
        # Accept common currency formats: $#,##0.00, #,##0.00, etc.
        currency_formats = ['$#,##0.00', '#,##0.00', '€#,##0.00', '£#,##0.00',
                            '_($#,##0.00', '"$"#,##0.00', '[$€-407]#,##0.00']
        g2_currency_ok = any(g2_format.startswith(fmt.rstrip('0').rstrip('.')) or
                              g2_format == fmt or '$' in g2_format or '€' in g2_format
                              for fmt in currency_formats)
        # Simpler check: format contains currency symbol or standard patterns
        g2_currency_ok = ('$' in g2_format or '€' in g2_format or '£' in g2_format or
                          '#,##0' in g2_format)

        if g2_bold and g2_currency_ok:
            print(f"PASS: Component 3 — G2 bold={g2_bold}, currency format='{g2_format}' (0.10 pts)")
            total_score += 0.10
        elif g2_bold:
            print(f"PARTIAL: Component 3 — G2 bold but format='{g2_format}' not currency (0.05 pts)")
            total_score += 0.05
        elif g2_currency_ok:
            print(f"PARTIAL: Component 3 — G2 has currency format but not bold (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 3 — G2 bold={g2_bold}, format='{g2_format}'")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: D2:D40 currency formatted (0.10 points)
    try:
        currency_formatted_count = 0
        for row in range(2, 41):
            d_format = ws.cell(row=row, column=4).number_format
            if ('$' in d_format or '€' in d_format or '£' in d_format or
                    '#,##0' in d_format):
                currency_formatted_count += 1

        if currency_formatted_count >= 35:
            print(f"PASS: Component 4 — D2:D40 currency formatted ({currency_formatted_count}/39 cells) (0.10 pts)")
            total_score += 0.10
        elif currency_formatted_count > 0:
            print(f"PARTIAL: Component 4 — Only {currency_formatted_count}/39 D cells currency formatted (0.04 pts)")
            total_score += 0.04
        else:
            print(f"FAIL: Component 4 — D2:D40 not currency formatted (count={currency_formatted_count})")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Table sorted by C column ascending (0.15 points)
    try:
        dates = []
        for row in range(2, 41):
            d = ws.cell(row=row, column=3).value
            if d is not None:
                dates.append(d)

        if len(dates) >= 2:
            is_sorted = all(dates[i] <= dates[i+1] for i in range(len(dates)-1))
            if is_sorted:
                print(f"PASS: Component 5 — Table sorted by Due Date ascending ({len(dates)} rows) (0.15 pts)")
                total_score += 0.15
            else:
                # Count out-of-order pairs
                out_of_order = sum(1 for i in range(len(dates)-1) if dates[i] > dates[i+1])
                print(f"FAIL: Component 5 — Table not sorted, {out_of_order} out-of-order pairs found")
        else:
            print(f"FAIL: Component 5 — Not enough date data (found {len(dates)} dates)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Conditional formatting on C2:C40 (0.20 points)
    # Expect: red rule for <=7 days (Pending), orange rule for <=14 days (Pending)
    try:
        cf_rules = ws.conditional_formatting
        cf_list = list(cf_rules)

        # Find CF rules on C2:C40
        c_range_rules = []
        for cf in cf_list:
            cf_str = str(cf)
            if 'C2' in cf_str and 'C40' in cf_str:
                c_range_rules.extend(cf.rules)

        red_rule_ok = False
        orange_rule_ok = False

        for rule in c_range_rules:
            # Check fill color
            fill_color = None
            try:
                if rule.dxf and rule.dxf.fill:
                    fill_color = rule.dxf.fill.fgColor.rgb
            except Exception:
                pass

            # Check formula content for 7-day and 14-day references
            formula_str = ''
            if hasattr(rule, 'formula') and rule.formula:
                formula_str = ' '.join(str(f) for f in rule.formula).upper()

            # Red rule: formula with +7 and "Pending"
            if fill_color and ('FF0000' in fill_color.upper() or fill_color.upper() == 'FFFF0000'):
                if '+7' in formula_str and 'PENDING' in formula_str:
                    red_rule_ok = True

            # Orange rule: formula with +14 and "Pending"
            if fill_color and ('FFA500' in fill_color.upper() or 'FF8000' in fill_color.upper() or
                               'FFC000' in fill_color.upper() or 'FFFFA500' in fill_color.upper()):
                if '+14' in formula_str and 'PENDING' in formula_str:
                    orange_rule_ok = True

        if red_rule_ok and orange_rule_ok:
            print(f"PASS: Component 6 — CF on C2:C40: red (7-day) and orange (14-day) rules found (0.20 pts)")
            total_score += 0.20
        elif red_rule_ok or orange_rule_ok:
            found = 'red (7-day)' if red_rule_ok else 'orange (14-day)'
            missing = 'orange (14-day)' if red_rule_ok else 'red (7-day)'
            print(f"PARTIAL: Component 6 — CF: {found} found, {missing} missing (0.10 pts)")
            total_score += 0.10
        elif len(c_range_rules) > 0:
            print(f"PARTIAL: Component 6 — CF rules found on C range but colors/formulas don't match expected (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 — No CF rules found on C2:C40")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
