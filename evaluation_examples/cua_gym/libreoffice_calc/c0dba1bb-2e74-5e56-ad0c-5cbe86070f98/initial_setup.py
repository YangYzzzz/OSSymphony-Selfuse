"""
Initial Setup: Create spreadsheet with numeric data in column B for dynamic named range task
Task ID: calc_nrv_034
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_034'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SensorReadings"

    # Column A: Timestamp labels, Column B: Values (numeric data)
    ws.cell(row=1, column=1, value="Timestamp")
    ws.cell(row=1, column=2, value="Values")

    # Header styling
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    for col in [1, 2]:
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Generate 119 rows of realistic sensor data (B2:B120)
    random.seed(42)
    base_date = "2025-01-01"
    from datetime import datetime, timedelta
    start_dt = datetime(2025, 1, 1, 8, 0, 0)

    for i in range(119):
        row = i + 2
        # Timestamp in column A
        ts = start_dt + timedelta(hours=i * 2)
        ws.cell(row=row, column=1, value=ts.strftime("%Y-%m-%d %H:%M"))
        # Numeric sensor reading in column B - varies between 15.0 and 85.0
        reading = round(random.uniform(15.0, 85.0), 2)
        ws.cell(row=row, column=2, value=reading)

    # Set column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14

    # Add a second sheet with some metadata
    ws2 = wb.create_sheet("Config")
    ws2["A1"] = "Sensor ID"
    ws2["B1"] = "TMP-4072"
    ws2["A2"] = "Location"
    ws2["B2"] = "Building C, Floor 3"
    ws2["A3"] = "Calibration Date"
    ws2["B3"] = "2025-01-01"
    ws2["A4"] = "Sampling Interval"
    ws2["B4"] = "2 hours"
    ws2["A5"] = "Unit"
    ws2["B5"] = "Celsius"
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 22

    # NO named ranges in initial state
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
