"""
Reward Script: Merge title, format, apply thick outer borders, alternating row fill
Task ID: calc_ggf_041
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): A1:G1 merged with correct title text
  Component 2 (0.20): Title formatting (bold 14pt, centered)
  Component 3 (0.30): Thick outer borders on A2:G30
  Component 4 (0.25): Alternating light gray fill on odd rows 3-29
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_041'


def persist_app_state(domain):
    """Try to save any unsaved LibreOffice edits via Ctrl+S."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Expenses' sheet must exist
    if 'Expenses' not in wb.sheetnames:
        print("CRITICAL: 'Expenses' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Expenses']

    # =====================================================================
    # Component 1: A1:G1 merged with correct title text (0.25 points)
    # =====================================================================
    try:
        # Check that A1:G1 is a merged range
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        has_merge = any('A1:G1' in mr for mr in merged_ranges)

        # Check title text
        a1_val = ws['A1'].value
        expected_title = 'Monthly Expense Report \u2013 Finance Department'
        has_title = (a1_val is not None and
                     str(a1_val).strip() == expected_title)

        if has_merge and has_title:
            print(f"PASS: Component 1 - A1:G1 merged with correct title (0.25 pts)")
            total_score += 0.25
        elif has_merge:
            print(f"FAIL: Component 1 - A1:G1 merged but title wrong: {repr(a1_val)}")
        elif has_title:
            print(f"FAIL: Component 1 - Title correct but A1:G1 not merged. Ranges: {merged_ranges}")
        else:
            print(f"FAIL: Component 1 - No merge and no title. A1={repr(a1_val)}, ranges={merged_ranges}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # =====================================================================
    # Component 2: Title formatting - bold 14pt, centered (0.20 points)
    # =====================================================================
    try:
        a1 = ws['A1']
        is_bold = (a1.font.bold is True)
        is_14pt = (a1.font.size is not None and abs(float(a1.font.size) - 14.0) < 0.5)
        is_centered = (a1.alignment.horizontal == 'center')

        sub_checks = 0
        if is_bold:
            sub_checks += 1
        if is_14pt:
            sub_checks += 1
        if is_centered:
            sub_checks += 1

        # Only award points if merge+title passed (prevents scoring on pre-existing empty A1)
        # Actually, we need to check that this is a task-introduced change.
        # In initial_env, A1 is empty with default font (11pt, not bold, general alignment).
        # These formatting properties only matter when the title is present (task change).
        # Gate: require the title text to be present for formatting to count.
        if a1.value is None or str(a1.value).strip() != expected_title:
            print(f"FAIL: Component 2 - Title not present, skipping format check")
        elif sub_checks == 3:
            print(f"PASS: Component 2 - Bold 14pt centered (0.20 pts)")
            total_score += 0.20
        else:
            earned = round(0.20 * sub_checks / 3, 2)
            total_score += earned
            details = f"bold={is_bold}, 14pt={is_14pt}, centered={is_centered}"
            print(f"PARTIAL: Component 2 - {sub_checks}/3 format checks ({earned} pts) [{details}]")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # =====================================================================
    # Component 3: Thick outer borders on A2:G30 (0.30 points)
    # =====================================================================
    try:
        border_pass = 0
        border_total = 0

        # Top edge: row 2, cols A-G should have thick top border
        for c in range(1, 8):
            border_total += 1
            if ws.cell(row=2, column=c).border.top.style == 'thick':
                border_pass += 1

        # Bottom edge: row 30, cols A-G should have thick bottom border
        for c in range(1, 8):
            border_total += 1
            if ws.cell(row=30, column=c).border.bottom.style == 'thick':
                border_pass += 1

        # Left edge: col A, rows 2-30 should have thick left border
        for r in range(2, 31):
            border_total += 1
            if ws.cell(row=r, column=1).border.left.style == 'thick':
                border_pass += 1

        # Right edge: col G, rows 2-30 should have thick right border
        for r in range(2, 31):
            border_total += 1
            if ws.cell(row=r, column=7).border.right.style == 'thick':
                border_pass += 1

        border_ratio = border_pass / border_total if border_total > 0 else 0
        # Require at least 90% of outer border cells to have thick borders
        if border_ratio >= 0.9:
            print(f"PASS: Component 3 - Thick outer borders ({border_pass}/{border_total} cells) (0.30 pts)")
            total_score += 0.30
        elif border_ratio >= 0.5:
            earned = round(0.30 * border_ratio, 2)
            total_score += earned
            print(f"PARTIAL: Component 3 - {border_pass}/{border_total} border cells correct ({earned} pts)")
        else:
            print(f"FAIL: Component 3 - Only {border_pass}/{border_total} border cells have thick borders")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # =====================================================================
    # Component 4: Alternating light gray fill on odd rows 3,5,...,29 (0.25 points)
    # =====================================================================
    try:
        odd_rows = list(range(3, 30, 2))  # 3,5,7,...,29
        fill_pass = 0
        fill_total = 0

        for r in odd_rows:
            for c in range(1, 8):  # All 7 columns
                fill_total += 1
                cell = ws.cell(row=r, column=c)
                try:
                    fill_type = cell.fill.patternType
                    fill_rgb = cell.fill.fgColor.rgb if cell.fill.fgColor else None
                except:
                    fill_type = None
                    fill_rgb = None

                # Accept any solid gray-ish fill (D9D9D9 is standard light gray,
                # but also accept similar shades like C0C0C0, BFBFBF, E0E0E0, etc.)
                if fill_type == 'solid' and fill_rgb is not None:
                    # Extract RGB from ARGB string
                    rgb_hex = fill_rgb[-6:]  # last 6 chars
                    try:
                        r_val = int(rgb_hex[0:2], 16)
                        g_val = int(rgb_hex[2:4], 16)
                        b_val = int(rgb_hex[4:6], 16)
                        # Light gray: R, G, B should be similar and > 160
                        is_gray = (abs(r_val - g_val) < 30 and
                                   abs(g_val - b_val) < 30 and
                                   r_val > 160)
                        if is_gray:
                            fill_pass += 1
                    except:
                        pass

        fill_ratio = fill_pass / fill_total if fill_total > 0 else 0

        # Also check that even rows do NOT have fill (to confirm alternating pattern)
        even_rows = list(range(4, 31, 2))  # 4,6,8,...,30
        even_no_fill = 0
        even_total = 0
        for r in even_rows:
            for c in range(1, 8):
                even_total += 1
                cell = ws.cell(row=r, column=c)
                try:
                    fill_type = cell.fill.patternType
                except:
                    fill_type = None
                if fill_type is None or fill_type == 'none':
                    even_no_fill += 1

        even_ratio = even_no_fill / even_total if even_total > 0 else 0

        if fill_ratio >= 0.9 and even_ratio >= 0.8:
            print(f"PASS: Component 4 - Alternating gray fill ({fill_pass}/{fill_total} odd cells filled, "
                  f"{even_no_fill}/{even_total} even cells clean) (0.25 pts)")
            total_score += 0.25
        elif fill_ratio >= 0.5:
            earned = round(0.25 * fill_ratio, 2)
            total_score += earned
            print(f"PARTIAL: Component 4 - {fill_pass}/{fill_total} odd cells filled ({earned} pts)")
        else:
            print(f"FAIL: Component 4 - Only {fill_pass}/{fill_total} odd row cells have gray fill")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
