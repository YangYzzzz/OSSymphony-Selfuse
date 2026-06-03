"""
Initial Setup: Create expense report spreadsheet with 7-column table, no formatting
Task ID: calc_ggf_041
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_041'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Expenses'

    # Row 1: A1:G1 left empty (agent will merge and add title)

    # Row 2: Headers
    headers = ['Date', 'Category', 'Description', 'Vendor', 'Amount', 'Approved By', 'Status']
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)

    # Rows 3-30: 28 rows of realistic expense data
    expense_data = [
        ['2025-03-01', 'Travel', 'Flight to Chicago for client meeting', 'United Airlines', 487.50, 'Sarah Chen', 'Approved'],
        ['2025-03-02', 'Office Supplies', 'Printer toner cartridges (x3)', 'Staples', 134.97, 'Sarah Chen', 'Approved'],
        ['2025-03-03', 'Meals', 'Team lunch - quarterly planning', 'Olive Garden', 215.80, 'Marcus Johnson', 'Approved'],
        ['2025-03-04', 'Software', 'Annual Tableau license renewal', 'Tableau Inc.', 1200.00, 'Sarah Chen', 'Pending'],
        ['2025-03-05', 'Travel', 'Hotel 2 nights - Chicago trip', 'Marriott Downtown', 378.00, 'Sarah Chen', 'Approved'],
        ['2025-03-06', 'Utilities', 'March electricity bill - Floor 3', 'ComEd', 892.45, 'David Park', 'Approved'],
        ['2025-03-07', 'Office Supplies', 'Ergonomic keyboards (x5)', 'Amazon Business', 449.95, 'Marcus Johnson', 'Approved'],
        ['2025-03-08', 'Training', 'CPA continuing education course', 'AICPA', 350.00, 'Sarah Chen', 'Approved'],
        ['2025-03-10', 'Meals', 'Client dinner - Acme Corp', 'Morton\'s Steakhouse', 312.60, 'David Park', 'Approved'],
        ['2025-03-11', 'Software', 'Slack premium - monthly', 'Slack Technologies', 156.00, 'Marcus Johnson', 'Approved'],
        ['2025-03-12', 'Travel', 'Uber rides - Chicago trip (3)', 'Uber', 87.30, 'Sarah Chen', 'Approved'],
        ['2025-03-13', 'Office Supplies', 'Copy paper - 10 reams', 'Office Depot', 64.90, 'Marcus Johnson', 'Approved'],
        ['2025-03-14', 'Maintenance', 'HVAC filter replacement', 'Johnson Controls', 275.00, 'David Park', 'Pending'],
        ['2025-03-15', 'Travel', 'Train tickets NYC-Boston roundtrip', 'Amtrak', 198.00, 'Sarah Chen', 'Approved'],
        ['2025-03-17', 'Meals', 'Department breakfast meeting', 'Panera Bread', 89.45, 'Marcus Johnson', 'Approved'],
        ['2025-03-18', 'Software', 'Adobe Creative Cloud - 2 seats', 'Adobe Systems', 109.98, 'Sarah Chen', 'Approved'],
        ['2025-03-19', 'Office Supplies', 'Standing desk converter', 'Varidesk', 395.00, 'David Park', 'Denied'],
        ['2025-03-20', 'Training', 'Excel advanced workshop (3 staff)', 'Coursera', 147.00, 'Marcus Johnson', 'Approved'],
        ['2025-03-21', 'Utilities', 'Internet service - March', 'AT&T Business', 249.99, 'David Park', 'Approved'],
        ['2025-03-22', 'Travel', 'Parking garage - monthly pass', 'SP Plus Corp', 185.00, 'Sarah Chen', 'Approved'],
        ['2025-03-24', 'Meals', 'Working lunch - audit prep', 'Chipotle', 67.80, 'Marcus Johnson', 'Approved'],
        ['2025-03-25', 'Maintenance', 'Office cleaning service - March', 'CleanPro LLC', 450.00, 'David Park', 'Approved'],
        ['2025-03-26', 'Software', 'Zoom enterprise - monthly', 'Zoom Video', 199.90, 'Sarah Chen', 'Approved'],
        ['2025-03-27', 'Office Supplies', 'Presentation binders and tabs', 'Staples', 42.15, 'Marcus Johnson', 'Approved'],
        ['2025-03-28', 'Travel', 'Mileage reimbursement - 230 miles', 'Employee Reimb.', 149.50, 'David Park', 'Pending'],
        ['2025-03-29', 'Training', 'Leadership seminar registration', 'AMA', 525.00, 'Sarah Chen', 'Approved'],
        ['2025-03-30', 'Meals', 'Month-end celebration lunch', 'The Capital Grille', 428.50, 'David Park', 'Approved'],
        ['2025-03-31', 'Utilities', 'Phone service - March', 'Verizon Business', 387.60, 'David Park', 'Approved'],
    ]

    for r, row_data in enumerate(expense_data, 3):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 38
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
