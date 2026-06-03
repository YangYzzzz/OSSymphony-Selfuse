"""
Initial Setup: Vendor evaluation matrix with raw scores across 6 criteria and 4 vendors.
Task ID: calc_ops_045
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_045'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'VendorEval'

    # --- Headers (Row 1) ---
    headers = ['Criterion', 'Weight', 'Vendor A', 'Vendor B', 'Vendor C', 'Vendor D']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # --- Criteria data (Rows 2-7) ---
    data = [
        ['Price Competitiveness', 0.25, 4, 5, 3, 4],
        ['Quality',               0.25, 5, 3, 4, 5],
        ['Delivery Reliability',  0.20, 3, 4, 5, 3],
        ['Technical Support',     0.10, 4, 3, 4, 2],
        ['Financial Stability',   0.10, 5, 4, 3, 4],
        ['Innovation',            0.10, 3, 4, 5, 3],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if c == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            if c == 2:
                cell.number_format = '0.00'

    # --- Row 8: empty separator ---

    # --- Row 9: Weighted Total label (NO formulas in C9:F9) ---
    label9 = ws.cell(row=9, column=1, value='Weighted Total')
    label9.font = Font(name='Calibri', size=11, bold=True)
    label9.border = border
    for c in range(2, 7):
        cell = ws.cell(row=9, column=c)
        cell.border = border

    # --- Row 10: Rank label (NO formulas in C10:F10) ---
    label10 = ws.cell(row=10, column=1, value='Rank')
    label10.font = Font(name='Calibri', size=11, bold=True)
    label10.border = border
    for c in range(2, 7):
        cell = ws.cell(row=10, column=c)
        cell.border = border

    # --- Column widths ---
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 10
    for col_letter in ['C', 'D', 'E', 'F']:
        ws.column_dimensions[col_letter].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
