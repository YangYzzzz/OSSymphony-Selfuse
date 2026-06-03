"""
Reward Script: Vendor evaluation matrix with weighted scoring, ranking, and conditional formatting
Task ID: calc_ops_045
Domain: libreoffice_calc
Scoring:
  Component 1 (0.40): SUMPRODUCT formulas in C9:F9
  Component 2 (0.35): RANK formulas in C10:F10
  Component 3 (0.25): Green fill on the rank-1 cell in row 10
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_045'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: VendorEval sheet must exist
    if 'VendorEval' not in wb.sheetnames:
        print("FAIL: Sheet 'VendorEval' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['VendorEval']

    # =========================================================================
    # Component 1: SUMPRODUCT formulas in C9:F9 (0.40 points)
    # Each of the 4 cells gets 0.10 points
    # =========================================================================
    try:
        sumproduct_count = 0
        for col_letter in ['C', 'D', 'E', 'F']:
            cell_ref = f'{col_letter}9'
            val = ws[cell_ref].value
            if val is not None and isinstance(val, str):
                # Normalize: remove spaces, uppercase
                normalized = val.upper().replace(" ", "")
                if 'SUMPRODUCT' in normalized:
                    print(f"PASS: {cell_ref} contains SUMPRODUCT formula: {val}")
                    sumproduct_count += 1
                else:
                    print(f"FAIL: {cell_ref} expected SUMPRODUCT formula, found: {val}")
            else:
                print(f"FAIL: {cell_ref} is empty or not a formula (value={val!r})")

        comp1_score = sumproduct_count * 0.10
        if comp1_score > 0:
            total_score += comp1_score
        print(f"Component 1 subtotal: {comp1_score}/0.40 ({sumproduct_count}/4 SUMPRODUCT formulas)")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # =========================================================================
    # Component 2: RANK formulas in C10:F10 (0.35 points)
    # Each of the 4 cells gets 0.0875 points
    # =========================================================================
    try:
        rank_count = 0
        for col_letter in ['C', 'D', 'E', 'F']:
            cell_ref = f'{col_letter}10'
            val = ws[cell_ref].value
            if val is not None and isinstance(val, str):
                normalized = val.upper().replace(" ", "")
                if 'RANK' in normalized:
                    print(f"PASS: {cell_ref} contains RANK formula: {val}")
                    rank_count += 1
                else:
                    print(f"FAIL: {cell_ref} expected RANK formula, found: {val}")
            else:
                print(f"FAIL: {cell_ref} is empty or not a formula (value={val!r})")

        comp2_score = rank_count * 0.0875
        if comp2_score > 0:
            total_score += comp2_score
        print(f"Component 2 subtotal: {comp2_score}/0.35 ({rank_count}/4 RANK formulas)")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # =========================================================================
    # Component 3: Green fill on the rank-1 cell (0.25 points)
    # The task says apply green fill to the cell with rank 1.
    # Ground truth: C10 is rank 1 (Vendor A has highest weighted total 4.05).
    # We check if ANY cell in C10:F10 has a green-ish fill. The cell that has
    # green fill should be the one whose SUMPRODUCT in row 9 is highest,
    # i.e. Vendor A (column C).
    # =========================================================================
    try:
        green_cells = []

        for col_letter in ['C', 'D', 'E', 'F']:
            cell_ref = f'{col_letter}10'
            cell = ws[cell_ref]
            fill_type = cell.fill.fill_type
            try:
                fg_rgb = cell.fill.fgColor.rgb if cell.fill.fgColor else None
            except Exception:
                fg_rgb = None

            if fill_type == 'solid' and fg_rgb is not None:
                # Check for green-ish colors (common greens in hex ARGB)
                # FF00B050 = standard green, FF00FF00 = bright green,
                # FF92D050 = light green, etc.
                rgb_str = str(fg_rgb).upper()
                # Extract RGB components (skip alpha prefix)
                if len(rgb_str) >= 6:
                    # Get last 6 chars as RGB
                    r_hex = rgb_str[-6:-4]
                    g_hex = rgb_str[-4:-2]
                    b_hex = rgb_str[-2:]
                    try:
                        r_val = int(r_hex, 16)
                        g_val = int(g_hex, 16)
                        b_val = int(b_hex, 16)
                        # Green: G channel significantly higher than R and B
                        if g_val > 100 and g_val > r_val and g_val > b_val:
                            green_cells.append(cell_ref)
                            print(f"PASS: {cell_ref} has green fill (ARGB={rgb_str})")
                    except ValueError:
                        pass

        if len(green_cells) > 0:
            # Additional check: green fill should be on C10 (the rank 1 vendor)
            if 'C10' in green_cells:
                print(f"PASS: Green fill correctly applied to C10 (rank 1 cell) ({0.25} pts)")
                total_score += 0.25
            elif len(green_cells) > 0:
                # Green fill exists but on wrong cell - partial credit
                print(f"PARTIAL: Green fill found on {green_cells} but expected on C10 (rank 1)")
                total_score += 0.15
        else:
            print("FAIL: No green fill detected on any cell in C10:F10")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
