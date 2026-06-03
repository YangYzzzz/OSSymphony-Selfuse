"""
Initial Setup: Create timesheet workbook with Summary pivot table for chart creation task
Task ID: calc_gg5_041
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_041'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Timesheet (raw data) ---
    ws_raw = wb.active
    ws_raw.title = "Timesheet"

    raw_headers = ["Date", "Employee", "Project", "Hours", "Task Type", "Notes"]
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    white_font = Font(bold=True, size=11, color="FFFFFF")

    for c, h in enumerate(raw_headers, 1):
        cell = ws_raw.cell(row=1, column=c, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    raw_data = [
        ["2025-03-03", "Sarah Chen", "Website Redesign", 7.5, "Development", "Frontend components"],
        ["2025-03-03", "Marcus Johnson", "Q1 Marketing", 8.0, "Planning", "Campaign strategy"],
        ["2025-03-03", "Priya Patel", "Website Redesign", 6.5, "Design", "UI mockups"],
        ["2025-03-04", "Sarah Chen", "Website Redesign", 8.0, "Development", "API integration"],
        ["2025-03-04", "James O'Brien", "Infrastructure", 7.0, "DevOps", "CI/CD pipeline"],
        ["2025-03-04", "Lisa Wang", "Client Portal", 8.5, "Development", "Auth module"],
        ["2025-03-05", "Marcus Johnson", "Q1 Marketing", 6.0, "Content", "Blog articles"],
        ["2025-03-05", "Aisha Mohammed", "Data Migration", 9.0, "Analysis", "Schema mapping"],
        ["2025-03-05", "Carlos Rivera", "Mobile App", 7.5, "Development", "Push notifications"],
        ["2025-03-06", "Sarah Chen", "Website Redesign", 8.0, "Testing", "Unit tests"],
        ["2025-03-06", "Priya Patel", "Client Portal", 7.0, "Design", "Dashboard layout"],
        ["2025-03-06", "David Kim", "Infrastructure", 8.0, "DevOps", "Monitoring setup"],
        ["2025-03-07", "Lisa Wang", "Client Portal", 7.5, "Development", "Payment gateway"],
        ["2025-03-07", "James O'Brien", "Infrastructure", 6.5, "DevOps", "Load balancing"],
        ["2025-03-07", "Aisha Mohammed", "Data Migration", 8.0, "Development", "ETL scripts"],
        ["2025-03-10", "Carlos Rivera", "Mobile App", 8.0, "Development", "Offline mode"],
        ["2025-03-10", "David Kim", "Infrastructure", 7.0, "DevOps", "Security audit"],
        ["2025-03-10", "Rachel Foster", "Q1 Marketing", 6.5, "Design", "Social media assets"],
        ["2025-03-11", "Sarah Chen", "Website Redesign", 7.5, "Development", "Performance tuning"],
        ["2025-03-11", "Marcus Johnson", "Q1 Marketing", 8.0, "Planning", "Budget review"],
        ["2025-03-11", "Tom Henderson", "Client Portal", 7.0, "QA", "Regression tests"],
        ["2025-03-12", "Priya Patel", "Mobile App", 6.5, "Design", "App icons"],
        ["2025-03-12", "Lisa Wang", "Client Portal", 8.0, "Development", "Report generation"],
        ["2025-03-12", "James O'Brien", "Infrastructure", 7.5, "DevOps", "Backup strategy"],
        ["2025-03-13", "Aisha Mohammed", "Data Migration", 8.5, "Development", "Data validation"],
        ["2025-03-13", "Carlos Rivera", "Mobile App", 7.0, "Testing", "Integration tests"],
        ["2025-03-13", "David Kim", "Infrastructure", 8.0, "DevOps", "Container orchestration"],
        ["2025-03-14", "Rachel Foster", "Q1 Marketing", 7.5, "Content", "Newsletter design"],
        ["2025-03-14", "Tom Henderson", "Client Portal", 6.0, "QA", "Performance tests"],
        ["2025-03-14", "Sarah Chen", "Website Redesign", 8.5, "Development", "Launch prep"],
        ["2025-03-17", "Marcus Johnson", "Q1 Marketing", 7.0, "Reporting", "Campaign metrics"],
        ["2025-03-17", "Priya Patel", "Website Redesign", 8.0, "Design", "Final revisions"],
        ["2025-03-17", "Lisa Wang", "Client Portal", 7.0, "Development", "Bug fixes"],
        ["2025-03-18", "James O'Brien", "Infrastructure", 8.5, "DevOps", "Disaster recovery"],
        ["2025-03-18", "Aisha Mohammed", "Data Migration", 7.5, "Development", "Rollback procedures"],
        ["2025-03-18", "Carlos Rivera", "Mobile App", 8.0, "Development", "Tablet layout"],
        ["2025-03-19", "David Kim", "Infrastructure", 7.5, "DevOps", "DNS configuration"],
        ["2025-03-19", "Rachel Foster", "Q1 Marketing", 8.0, "Design", "Landing page"],
        ["2025-03-19", "Tom Henderson", "Client Portal", 7.5, "QA", "UAT coordination"],
        ["2025-03-20", "Sarah Chen", "Website Redesign", 7.0, "Documentation", "Tech specs"],
        ["2025-03-20", "Marcus Johnson", "Q1 Marketing", 6.5, "Reporting", "ROI analysis"],
        ["2025-03-20", "Priya Patel", "Client Portal", 7.5, "Design", "Style guide update"],
        ["2025-03-21", "Lisa Wang", "Mobile App", 8.0, "Development", "API endpoints"],
        ["2025-03-21", "Aisha Mohammed", "Data Migration", 8.0, "Testing", "Data integrity checks"],
        ["2025-03-21", "Tom Henderson", "Website Redesign", 6.5, "QA", "Cross-browser tests"],
    ]

    for r, row_data in enumerate(raw_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_raw.cell(row=r, column=c, value=val)

    # Set column widths
    ws_raw.column_dimensions["A"].width = 14
    ws_raw.column_dimensions["B"].width = 20
    ws_raw.column_dimensions["C"].width = 20
    ws_raw.column_dimensions["D"].width = 10
    ws_raw.column_dimensions["E"].width = 14
    ws_raw.column_dimensions["F"].width = 25

    # --- Sheet 2: Summary (pivot table in A3:F20) ---
    ws_summary = wb.create_sheet("Summary")

    # Title row
    ws_summary.merge_cells("A1:F1")
    title_cell = ws_summary["A1"]
    title_cell.value = "Monthly Timesheet Summary - March 2025"
    title_cell.font = Font(size=14, bold=True, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Subtitle row
    ws_summary.merge_cells("A2:F2")
    sub_cell = ws_summary["A2"]
    sub_cell.value = "Hours by Employee (All Projects)"
    sub_cell.font = Font(size=10, italic=True, color="666666")
    sub_cell.alignment = Alignment(horizontal="center")

    # Pivot table headers in row 3
    pivot_headers = ["Employee", "Week 1", "Week 2", "Week 3", "Week 4", "Total Hours"]
    pivot_header_fill = PatternFill(start_color="FF2E75B6", end_color="FF2E75B6", fill_type="solid")
    pivot_header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999"),
    )

    for c, h in enumerate(pivot_headers, 1):
        cell = ws_summary.cell(row=3, column=c, value=h)
        cell.font = pivot_header_font
        cell.fill = pivot_header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Employee data rows (A4:F20 = 17 employees)
    # Each employee has weekly hours and a total
    employees = [
        ["Sarah Chen",        7.5, 16.0, 7.5, 7.0, 38.0],
        ["Marcus Johnson",    8.0,  6.0, 8.0, 13.5, 35.5],
        ["Priya Patel",       6.5,  7.0, 14.5, 7.5, 35.5],
        ["James O'Brien",     7.0,  6.5, 7.5, 8.5, 29.5],
        ["Lisa Wang",         8.5,  7.5, 15.0, 8.0, 39.0],
        ["Aisha Mohammed",    9.0,  8.0, 8.5, 15.5, 41.0],
        ["Carlos Rivera",     7.5,  8.0, 7.0, 8.0, 30.5],
        ["David Kim",         8.0,  7.0, 8.0, 7.5, 30.5],
        ["Rachel Foster",     0.0,  6.5, 7.5, 8.0, 22.0],
        ["Tom Henderson",     0.0,  7.0, 6.0, 14.0, 27.0],
        ["Elena Vasquez",     6.0,  7.5, 8.0, 7.0, 28.5],
        ["Ryan Mitchell",     7.0,  8.5, 6.5, 8.0, 30.0],
        ["Nadia Kowalski",    8.0,  7.0, 7.5, 6.5, 29.0],
        ["Kevin Okafor",      7.5,  6.0, 8.5, 7.0, 29.0],
        ["Maya Singh",        6.5,  8.0, 7.0, 8.5, 30.0],
        ["Brandon Lee",       8.0,  7.5, 6.0, 7.0, 28.5],
        ["Olivia Thompson",   7.0,  6.5, 8.0, 7.5, 29.0],
    ]

    even_fill = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")

    for r, row_data in enumerate(employees, 4):
        for c, val in enumerate(row_data, 1):
            cell = ws_summary.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if r % 2 == 0:
                cell.fill = even_fill
            if c == 1:
                cell.font = Font(bold=True)
            elif c >= 2:
                cell.number_format = '0.0'
                cell.alignment = Alignment(horizontal="center")

    # Set column widths for Summary
    ws_summary.column_dimensions["A"].width = 22
    ws_summary.column_dimensions["B"].width = 12
    ws_summary.column_dimensions["C"].width = 12
    ws_summary.column_dimensions["D"].width = 12
    ws_summary.column_dimensions["E"].width = 12
    ws_summary.column_dimensions["F"].width = 14

    # Row height for header
    ws_summary.row_dimensions[1].height = 28
    ws_summary.row_dimensions[3].height = 22

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
