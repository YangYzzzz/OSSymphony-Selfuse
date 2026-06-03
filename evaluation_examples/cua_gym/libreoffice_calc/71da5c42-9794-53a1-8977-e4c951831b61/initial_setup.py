"""
Initial Setup: Energy consumption table with 4 energy types and 5 years of data (no totals/averages/charts)
Task ID: osworld_calc_total_row_line_chart_006
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_total_row_line_chart_006'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Energy Consumption ---
    ws = wb.active
    ws.title = 'Energy Consumption'

    # Headers
    years = [2019, 2020, 2021, 2022, 2023]
    headers = ['Energy Type'] + years
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Energy consumption data (realistic values in MWh or similar unit)
    data = [
        ['Electricity', 45230, 47580, 46910, 49340, 51200],
        ['Gas',         28640, 27950, 29100, 28430, 27680],
        ['Solar',        3820,  4150,  5270,  6380,  7940],
        ['Wind',         2110,  2480,  3020,  3750,  4620],
    ]

    alt_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    for r_idx, row_data in enumerate(data, 2):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx % 2 == 0:
                cell.fill = alt_fill
            if c_idx > 1:
                cell.alignment = Alignment(horizontal='right')

    # Column widths
    ws.column_dimensions['A'].width = 16
    for col_idx in range(2, 7):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
