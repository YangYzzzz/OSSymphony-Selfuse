"""
Reward Script: Add total and average rows, then create a line chart of annual totals
Task ID: osworld_calc_total_row_line_chart_006
Domain: libreoffice_calc
Scoring:
  Component 1: Total row at row 6 with 'Total' label and SUM formulas (0.35 pts)
  Component 2: Average row at row 7 with 'Average' label and AVERAGE formulas (0.25 pts)
  Component 3: A LineChart exists on the sheet (0.20 pts)
  Component 4: LineChart has 4 series matching energy type names (0.20 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_total_row_line_chart_006'

SHEET_NAME = 'Energy Consumption'
ENERGY_TYPES = ['Electricity', 'Gas', 'Solar', 'Wind']
YEARS = [2019, 2020, 2021, 2022, 2023]  # columns B-F (cols 2-6)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook — precondition gate
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet exists — precondition gate
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # --------------------------------------------------------------------------
    # Component 1: Total row at row 6 with 'Total' label and SUM formulas (0.35 pts)
    # The initial file has only 5 rows (header + 4 energy types).
    # Adding the Total row is the primary task-introduced change.
    # --------------------------------------------------------------------------
    try:
        total_label = ws.cell(row=6, column=1).value
        # Check label (case-insensitive)
        label_ok = total_label is not None and str(total_label).strip().lower() == 'total'

        # Check SUM formulas in columns B-F (cols 2-6) for row 6
        # The formula should reference rows 2-5 (the 4 energy type rows)
        sum_formulas_ok = 0
        for col in range(2, 7):  # columns B through F
            cell_val = ws.cell(row=6, column=col).value
            if cell_val is not None:
                cell_str = str(cell_val).strip().upper().replace(' ', '')
                # Accept any SUM formula referencing a range ending at row 5
                if cell_str.startswith('=SUM(') and '5' in cell_str:
                    sum_formulas_ok += 1

        if label_ok and sum_formulas_ok >= 5:
            print(f"PASS: Component 1 — Total row at row 6: label='{total_label}', "
                  f"SUM formulas in {sum_formulas_ok}/5 year columns (0.35 pts)")
            total_score += 0.35
        elif label_ok and sum_formulas_ok >= 3:
            # Partial: label correct and most formulas present
            print(f"PASS (partial): Component 1 — Total row label correct but only "
                  f"{sum_formulas_ok}/5 SUM formulas found (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — Total row at row 6. "
                  f"Label: '{total_label}' (expected 'Total'), "
                  f"SUM formulas found: {sum_formulas_ok}/5")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # --------------------------------------------------------------------------
    # Component 2: Average row at row 7 with 'Average' label and AVERAGE formulas (0.25 pts)
    # --------------------------------------------------------------------------
    try:
        avg_label = ws.cell(row=7, column=1).value
        # Check label (case-insensitive)
        label_ok = avg_label is not None and str(avg_label).strip().lower() == 'average'

        # Check AVERAGE formulas in columns B-F (cols 2-6) for row 7
        avg_formulas_ok = 0
        for col in range(2, 7):  # columns B through F
            cell_val = ws.cell(row=7, column=col).value
            if cell_val is not None:
                cell_str = str(cell_val).strip().upper().replace(' ', '')
                # Accept any AVERAGE formula referencing a range ending at row 5
                if cell_str.startswith('=AVERAGE(') and '5' in cell_str:
                    avg_formulas_ok += 1

        if label_ok and avg_formulas_ok >= 5:
            print(f"PASS: Component 2 — Average row at row 7: label='{avg_label}', "
                  f"AVERAGE formulas in {avg_formulas_ok}/5 year columns (0.25 pts)")
            total_score += 0.25
        elif label_ok and avg_formulas_ok >= 3:
            print(f"PASS (partial): Component 2 — Average row label correct but only "
                  f"{avg_formulas_ok}/5 AVERAGE formulas found (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 — Average row at row 7. "
                  f"Label: '{avg_label}' (expected 'Average'), "
                  f"AVERAGE formulas found: {avg_formulas_ok}/5")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # --------------------------------------------------------------------------
    # Component 3: A LineChart exists on the sheet (0.20 pts)
    # The initial file has 0 charts; golden file adds a line chart.
    # --------------------------------------------------------------------------
    try:
        charts = ws._charts
        line_charts = [c for c in charts if type(c).__name__ == 'LineChart']

        if len(line_charts) >= 1:
            print(f"PASS: Component 3 — LineChart found on sheet "
                  f"({len(line_charts)} line chart(s) total) (0.20 pts)")
            total_score += 0.20
        elif len(charts) >= 1:
            # Chart exists but wrong type
            chart_types = [type(c).__name__ for c in charts]
            print(f"FAIL: Component 3 — Chart exists but is not a LineChart. "
                  f"Found: {chart_types}")
        else:
            print("FAIL: Component 3 — No charts found on sheet (expected 1 LineChart)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # --------------------------------------------------------------------------
    # Component 4: LineChart has series for each energy type (0.20 pts)
    # The chart should have 4 series: Electricity, Gas, Solar, Wind.
    # The series titles should match the energy type names.
    # --------------------------------------------------------------------------
    try:
        charts = ws._charts
        line_charts = [c for c in charts if type(c).__name__ == 'LineChart']

        if not line_charts:
            print("FAIL: Component 4 — No LineChart found, cannot check series")
        else:
            chart = line_charts[0]
            series_titles = []
            for s in chart.series:
                if s.title is not None:
                    # Series title can be a SeriesLabel object with attribute v
                    if hasattr(s.title, 'v') and s.title.v:
                        series_titles.append(str(s.title.v).strip())
                    elif hasattr(s.title, 'strRef') and s.title.strRef:
                        series_titles.append(str(s.title.strRef))
                    else:
                        series_titles.append(str(s.title))

            num_series = len(chart.series)
            # Check for at least 4 series
            if num_series >= 4:
                # Check if series titles match energy types (case-insensitive)
                expected_lower = {e.lower() for e in ENERGY_TYPES}
                found_lower = {t.lower() for t in series_titles}
                matched = expected_lower.intersection(found_lower)
                if len(matched) >= 4:
                    print(f"PASS: Component 4 — LineChart has {num_series} series "
                          f"with energy type labels: {series_titles} (0.20 pts)")
                    total_score += 0.20
                elif len(matched) >= 2:
                    print(f"PASS (partial): Component 4 — LineChart has {num_series} series "
                          f"but only {len(matched)}/4 energy type labels match. "
                          f"Found: {series_titles} (0.10 pts)")
                    total_score += 0.10
                else:
                    print(f"FAIL: Component 4 — LineChart has {num_series} series "
                          f"but energy type labels not found. Series: {series_titles}")
            else:
                print(f"FAIL: Component 4 — LineChart has only {num_series} series "
                      f"(expected 4 for Electricity/Gas/Solar/Wind)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
