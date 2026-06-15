"""
Initial Setup: Student test score table (no total row, no chart)
Task ID: osworld_calc_total_row_line_chart_005
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_total_row_line_chart_005'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Student Scores ---
    ws = wb.active
    ws.title = 'Student Scores'

    # Headers
    headers = ['Student Name', 'Round 1', 'Round 2', 'Round 3', 'Round 4', 'Round 5']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, name='Calibri', size=11, color='FFFFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Realistic student data (10 students, 5 rounds)
    students = [
        ['Emma Hartwell',    78,  82,  80,  85,  88],
        ['Liam Nguyen',      65,  70,  68,  72,  75],
        ['Sofia Ramirez',    90,  88,  92,  95,  91],
        ['Noah Patel',       55,  60,  58,  63,  67],
        ['Ava Johansson',    83,  85,  87,  84,  89],
        ['Marcus Williams',  71,  74,  76,  79,  81],
        ['Chloe Fischer',    88,  90,  86,  92,  94],
        ['Ethan Brooks',     60,  63,  65,  68,  70],
        ['Isabella Torres',  95,  93,  97,  96,  98],
        ['James O\'Brien',   72,  75,  77,  80,  78],
    ]

    for r, row_data in enumerate(students, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(horizontal='center' if c > 1 else 'left')

    # Column widths
    ws.column_dimensions['A'].width = 20
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col_letter].width = 10

    # Row height for header
    ws.row_dimensions[1].height = 20

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
