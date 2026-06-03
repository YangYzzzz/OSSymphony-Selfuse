"""
Reward Script: Add total row with SUM formulas and line chart showing student score trends
Task ID: osworld_calc_total_row_line_chart_005
Domain: libreoffice_calc
Scoring:
  Component 1: Total row present with SUM formulas in columns B-F (0.4 pts)
  Component 2: Average row present with AVERAGE formulas in columns B-F (0.2 pts)
  Component 3: A line chart exists in the spreadsheet (0.2 pts)
  Component 4: The line chart has a title containing the expected text (0.2 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_total_row_line_chart_005'

EXPECTED_CHART_TITLE = 'Student Score Trends Across Test Rounds'


def extract_chart_title(title_obj):
    """Extract plain-text title from openpyxl chart Title object."""
    if title_obj is None:
        return None
    try:
        tx = title_obj.tx
        if tx and tx.rich:
            paragraphs = tx.rich.p
            text_parts = []
            for p in paragraphs:
                for r in p.r:
                    text_parts.append(r.t)
            return ''.join(text_parts)
    except Exception:
        pass
    # Fallback: try string representation
    try:
        return str(title_obj)
    except Exception:
        return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Component 1: Total row with SUM formulas in B-F (0.4 points)
    # The task requires a Total row be appended to the student data table.
    # Golden state has row 12 as 'Total' with =SUM(B2:B11) etc. in columns B-F.
    # We search for any row whose column A cell is 'Total' (case-insensitive) and
    # columns B-F each contain a SUM formula.
    try:
        total_row_found = False
        for row_idx in range(1, ws.max_row + 1):
            label = ws.cell(row=row_idx, column=1).value
            if label is not None and str(label).strip().lower() == 'total':
                # Check that columns B-F all have SUM formulas
                sum_cols_ok = True
                for col_idx in range(2, 7):  # columns B to F (2 to 6)
                    cell_val = ws.cell(row=row_idx, column=col_idx).value
                    if cell_val is None or not isinstance(cell_val, str):
                        sum_cols_ok = False
                        break
                    if 'SUM' not in cell_val.upper():
                        sum_cols_ok = False
                        break
                if sum_cols_ok:
                    total_row_found = True
                    print(f"PASS: Component 1 — Total row found at row {row_idx} with SUM formulas in B-F (0.4 pts)")
                    total_score += 0.4
                    break
                else:
                    print(f"FAIL: Component 1 — Row with label 'Total' found at row {row_idx} but SUM formulas missing in B-F")
                    break
        if not total_row_found and total_score < 0.4:
            # Check if Total row exists but without proper formulas (no break happened)
            found_label = any(
                ws.cell(row=r, column=1).value is not None and
                str(ws.cell(row=r, column=1).value).strip().lower() == 'total'
                for r in range(1, ws.max_row + 1)
            )
            if not found_label:
                print("FAIL: Component 1 — No row with label 'Total' found in the spreadsheet")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Average row with AVERAGE formulas in B-F (0.2 points)
    # The context specifies an Average row should also be appended.
    try:
        avg_row_found = False
        for row_idx in range(1, ws.max_row + 1):
            label = ws.cell(row=row_idx, column=1).value
            if label is not None and str(label).strip().lower() == 'average':
                # Check that columns B-F all have AVERAGE formulas
                avg_cols_ok = True
                for col_idx in range(2, 7):
                    cell_val = ws.cell(row=row_idx, column=col_idx).value
                    if cell_val is None or not isinstance(cell_val, str):
                        avg_cols_ok = False
                        break
                    if 'AVERAGE' not in cell_val.upper():
                        avg_cols_ok = False
                        break
                if avg_cols_ok:
                    avg_row_found = True
                    print(f"PASS: Component 2 — Average row found at row {row_idx} with AVERAGE formulas in B-F (0.2 pts)")
                    total_score += 0.2
                    break
                else:
                    print(f"FAIL: Component 2 — Row with label 'Average' found at row {row_idx} but AVERAGE formulas missing in B-F")
                    break
        if not avg_row_found and total_score < 0.6:
            found_label = any(
                ws.cell(row=r, column=1).value is not None and
                str(ws.cell(row=r, column=1).value).strip().lower() == 'average'
                for r in range(1, ws.max_row + 1)
            )
            if not found_label:
                print("FAIL: Component 2 — No row with label 'Average' found in the spreadsheet")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: A line chart exists in the spreadsheet (0.2 points)
    # The task requires a line chart showing per-student score progression.
    try:
        charts = ws._charts
        line_chart_found = False
        line_chart_index = None
        for idx, chart in enumerate(charts):
            if type(chart).__name__ == 'LineChart':
                line_chart_found = True
                line_chart_index = idx
                break
        if line_chart_found:
            print(f"PASS: Component 3 — Line chart found (index {line_chart_index}) (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — No LineChart found in the worksheet. Charts present: {[type(c).__name__ for c in charts]}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Line chart has correct title (0.2 points)
    # The context specifies chart title = 'Student Score Trends Across Test Rounds'
    try:
        charts = ws._charts
        if line_chart_index is not None:
            chart = charts[line_chart_index]
            actual_title = extract_chart_title(chart.title)
            if actual_title and EXPECTED_CHART_TITLE.lower() in actual_title.lower():
                print(f"PASS: Component 4 — Chart title matches '{EXPECTED_CHART_TITLE}' (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 4 — Expected chart title '{EXPECTED_CHART_TITLE}', got '{actual_title}'")
        else:
            print("FAIL: Component 4 — Skipped (no line chart present)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
