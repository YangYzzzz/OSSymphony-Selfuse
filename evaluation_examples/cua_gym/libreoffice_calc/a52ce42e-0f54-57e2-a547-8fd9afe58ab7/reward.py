"""
Reward Script: Supply Chain PO Value — Open Commitment by Supplier
Task ID: calc_ops_supply_chain_po_value_008
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: POLines F2:F121 line value formulas (=Dx*Ex)       — 0.35 pts
  Component 2: SpendSummary B2:B8 SUMIFS formulas (by supplier, Open status) — 0.35 pts
  Component 3: SpendSummary total row (A9='Total', B9=SUM formula)  — 0.20 pts
  Component 4: Currency number format on value columns               — 0.10 pts
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_supply_chain_po_value_008'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: required sheets must exist
    required_sheets = ['POLines', 'SpendSummary']
    for sheet_name in required_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"CRITICAL: Required sheet '{sheet_name}' not found. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0

    ws_po = wb['POLines']
    ws_ss = wb['SpendSummary']

    # -------------------------------------------------------------------------
    # Component 1: POLines F2:F121 contain line value formulas =D{r}*E{r}
    # (0.35 points)
    # In the initial file, F2:F121 are all None. After the task, each cell
    # must contain a multiplication formula of the Qty * Unit Price columns.
    # -------------------------------------------------------------------------
    try:
        formula_count = 0
        correct_formula_count = 0
        expected_rows = range(2, 122)  # rows 2-121 inclusive

        for row in expected_rows:
            cell_val = ws_po.cell(row=row, column=6).value
            if cell_val is not None:
                formula_count += 1
                # Accept formulas like =D2*E2 or =E2*D2 (order-agnostic)
                val_str = str(cell_val).upper().replace(" ", "")
                expected_fwd = f"=D{row}*E{row}"
                expected_rev = f"=E{row}*D{row}"
                if val_str == expected_fwd.upper() or val_str == expected_rev.upper():
                    correct_formula_count += 1

        if correct_formula_count == 120:
            print(f"PASS: Component 1 — All 120 line value formulas present in POLines F2:F121 (0.35 pts)")
            total_score += 0.35
        elif correct_formula_count >= 60:
            # Partial: more than half correct
            partial = round(0.35 * (correct_formula_count / 120), 4)
            print(f"PARTIAL: Component 1 — {correct_formula_count}/120 line value formulas correct, "
                  f"awarded {partial} pts (out of 0.35)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {correct_formula_count}/120 correct line value formulas "
                  f"in POLines F column (found {formula_count} non-empty cells)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: SpendSummary B2:B8 contain SUMIFS formulas
    # Each formula must: sum POLines!F column, filter by supplier (col B),
    # AND filter by Status = "Open" (col G).
    # (0.35 points)
    # -------------------------------------------------------------------------
    try:
        sumifs_count = 0
        sumifs_rows = range(2, 9)  # rows 2-8 inclusive (7 suppliers)

        for row in sumifs_rows:
            cell_val = ws_ss.cell(row=row, column=2).value
            if cell_val is not None:
                val_str = str(cell_val).upper().replace(" ", "")
                # Must contain SUMIFS referencing POLines!F column and "Open" status
                has_sumifs = "SUMIFS" in val_str
                refs_polines_f = "POLINES!F" in val_str
                refs_open_status = '"OPEN"' in val_str or "'OPEN'" in val_str
                # Accept SUMIF (without S) as acceptable alternative
                has_sumif = "SUMIF" in val_str
                if (has_sumifs or has_sumif) and refs_polines_f and refs_open_status:
                    sumifs_count += 1

        if sumifs_count == 7:
            print(f"PASS: Component 2 — All 7 SUMIFS formulas present in SpendSummary B2:B8 (0.35 pts)")
            total_score += 0.35
        elif sumifs_count >= 4:
            partial = round(0.35 * (sumifs_count / 7), 4)
            print(f"PARTIAL: Component 2 — {sumifs_count}/7 SUMIFS formulas correct, "
                  f"awarded {partial} pts (out of 0.35)")
            total_score += partial
        else:
            # Check how many cells at least have some formula
            nonempty = sum(
                1 for r in sumifs_rows
                if ws_ss.cell(row=r, column=2).value is not None
            )
            print(f"FAIL: Component 2 — Only {sumifs_count}/7 valid SUMIFS formulas in SpendSummary B2:B8 "
                  f"({nonempty} non-empty cells total)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: SpendSummary total row — A9 = "Total", B9 = SUM formula
    # (0.20 points)
    # In the initial file, row 9 does not exist (max_row=8). After the task,
    # a total row must be added.
    # -------------------------------------------------------------------------
    try:
        a9_val = ws_ss.cell(row=9, column=1).value
        b9_val = ws_ss.cell(row=9, column=2).value

        a9_ok = a9_val is not None and str(a9_val).strip().lower() == "total"
        b9_ok = (b9_val is not None and
                 isinstance(b9_val, str) and
                 "SUM" in b9_val.upper() and
                 "B2" in b9_val.upper())

        if a9_ok and b9_ok:
            print(f"PASS: Component 3 — Total row found: A9={repr(a9_val)}, B9={repr(b9_val)} (0.20 pts)")
            total_score += 0.20
        elif a9_ok and not b9_ok:
            print(f"PARTIAL: Component 3 — 'Total' label present in A9 but B9 is missing/invalid "
                  f"(B9={repr(b9_val)}), awarded 0.10 pts")
            total_score += 0.10
        elif not a9_ok and b9_ok:
            print(f"PARTIAL: Component 3 — SUM formula in B9 but 'Total' label missing from A9 "
                  f"(A9={repr(a9_val)}), awarded 0.10 pts")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — Total row missing or incorrect: "
                  f"A9={repr(a9_val)}, B9={repr(b9_val)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: Currency number format applied to value columns
    # POLines F column and SpendSummary B column should have currency format
    # (e.g., '$#,##0.00' or similar currency format)
    # (0.10 points)
    # -------------------------------------------------------------------------
    try:
        currency_formats = {'$#,##0.00', '#,##0.00', '$#,##0', '"$"#,##0.00', '[$]#,##0.00'}

        # Check POLines F2 currency format
        po_f2_fmt = ws_po.cell(row=2, column=6).number_format
        po_currency = any(fmt in (po_f2_fmt or '') for fmt in currency_formats)

        # Check SpendSummary B2 currency format
        ss_b2_fmt = ws_ss.cell(row=2, column=2).number_format
        ss_currency = any(fmt in (ss_b2_fmt or '') for fmt in currency_formats)

        if po_currency and ss_currency:
            print(f"PASS: Component 4 — Currency format applied to both POLines F "
                  f"(format={repr(po_f2_fmt)}) and SpendSummary B (format={repr(ss_b2_fmt)}) (0.10 pts)")
            total_score += 0.10
        elif po_currency or ss_currency:
            print(f"PARTIAL: Component 4 — Currency format applied to only one column "
                  f"(POLines F format={repr(po_f2_fmt)}, SpendSummary B format={repr(ss_b2_fmt)}), "
                  f"awarded 0.05 pts")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4 — No currency format detected "
                  f"(POLines F format={repr(po_f2_fmt)}, SpendSummary B format={repr(ss_b2_fmt)})")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
