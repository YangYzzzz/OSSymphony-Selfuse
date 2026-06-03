"""
Initial Setup: Supply Chain PO Value Tracking
Task ID: calc_ops_supply_chain_po_value_008
Domain: libreoffice_calc

Creates a workbook with:
  - POLines sheet: 120 PO line items (columns A-E and G filled, F empty)
  - SpendSummary sheet: 7 suppliers in column A, column B empty
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_supply_chain_po_value_008'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    # -------------------------------------------------------
    # Sheet 1: POLines
    # -------------------------------------------------------
    ws = wb.active
    ws.title = 'POLines'

    # Headers
    headers = ['PO Number', 'Supplier', 'Item Description', 'Qty Ordered', 'Unit Price', 'Line Value', 'PO Status']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Suppliers
    suppliers = [
        'Acme Industrial Supplies',
        'Global Fasteners Co.',
        'Pacific Rim Electronics',
        'Midwest Metals Inc.',
        'Eastern Logistics Group',
        'Northern Chemical Corp.',
        'SouthWest Packaging Ltd.',
    ]

    # Item descriptions by category
    items = [
        'Stainless Steel Bolts M8x40', 'Hex Nuts M8', 'Flat Washers 8mm',
        'Electrical Cable 3-Core 2.5mm', 'Circuit Breaker 16A', 'Terminal Blocks 10-way',
        'Steel Sheet 4x8 3mm', 'Aluminium Extrusion 40x40 2m', 'Copper Pipe 22mm 3m',
        'Corrugated Cardboard Boxes L', 'Bubble Wrap Roll 500mm', 'Packing Tape 48mm',
        'Isopropyl Alcohol 5L', 'Machine Oil ISO 46 20L', 'Cutting Fluid 1L',
        'Safety Gloves Cut-5 Pair', 'Safety Glasses Clear Pair', 'Hard Hat Yellow',
        'Hydraulic Fitting 1/2 BSP', 'O-Ring Kit Metric Assorted', 'Silicone Sealant 310ml',
        'LED Strip Light 5m 24V', 'Power Supply 24V 5A', 'Pushbutton Green N/O',
        'Filter Cartridge 5-micron', 'Pump Seal Kit 50mm', 'Ball Valve 1/2 SS',
        'Angle Iron 40x40x3 3m', 'Mild Steel Rod 10mm 3m', 'Galv. Wire 1.6mm Coil',
    ]

    statuses = ['Open', 'Open', 'Open', 'Received', 'Cancelled', 'Open', 'Received', 'Open']

    import random
    random.seed(42)

    for i in range(120):
        row = i + 2
        po_num = f'PO-2025-{1000 + i:04d}'
        supplier = suppliers[i % len(suppliers)]
        item = items[i % len(items)]
        qty = random.choice([5, 10, 20, 25, 50, 100, 200, 500])
        unit_price = round(random.uniform(1.50, 250.00), 2)
        status = statuses[i % len(statuses)]

        ws.cell(row=row, column=1, value=po_num)
        ws.cell(row=row, column=2, value=supplier)
        ws.cell(row=row, column=3, value=item)
        ws.cell(row=row, column=4, value=qty)
        ws.cell(row=row, column=5, value=unit_price)
        # Column F (Line Value) intentionally left empty
        ws.cell(row=row, column=7, value=status)

    # Column widths for readability
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 12

    # -------------------------------------------------------
    # Sheet 2: SpendSummary
    # -------------------------------------------------------
    ws2 = wb.create_sheet('SpendSummary')

    # Headers
    ws2.cell(row=1, column=1, value='Supplier').font = Font(bold=True)
    ws2.cell(row=1, column=2, value='Open Commitment ($)').font = Font(bold=True)

    # 7 supplier names - column B intentionally empty
    for i, sup in enumerate(suppliers):
        ws2.cell(row=i + 2, column=1, value=sup)
        # Column B left empty

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 22

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: POLines (120 rows, F empty), SpendSummary (7 suppliers, B empty)')

create_initial()
