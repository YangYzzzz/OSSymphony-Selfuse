"""
Initial Setup: Sales capacity planning model - parameter sheet with raw inputs only
Task ID: calc_sales_072
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_072'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Capacity'

    # --- Headers ---
    ws['A1'] = 'Parameter'
    ws['B1'] = 'Value'

    # Style headers
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    for cell in [ws['A1'], ws['B1']]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Input parameters (rows 2-9) ---
    params = [
        ('Annual Revenue Target', 10000000),
        ('Avg Quota per Rep', 500000),
        ('Expected Attainment', 0.85),
        ('Effective Quota', None),          # Task: formula goes here
        ('Reps Needed (Productive)', None), # Task: formula goes here
        ('Annual Attrition Rate', 0.15),
        ('Avg Ramp Months', 6),
        ('Ramp Productivity', 0.50),
        ('Replacement Reps Needed', None),  # Task: formula goes here
        ('Ramp Adjustment', None),          # Task: formula goes here
        ('Total Headcount Needed', None),   # Task: formula goes here
    ]

    param_font = Font(name='Calibri', size=11)
    value_font = Font(name='Calibri', size=11)
    currency_fmt = '$#,##0'
    pct_fmt = '0%'
    int_fmt = '0'

    for i, (label, value) in enumerate(params, 2):
        ws.cell(row=i, column=1, value=label).font = param_font
        cell_b = ws.cell(row=i, column=2, value=value)
        cell_b.font = value_font

    # Number formats for input cells
    ws['B2'].number_format = currency_fmt   # Revenue Target
    ws['B3'].number_format = currency_fmt   # Quota per Rep
    ws['B4'].number_format = pct_fmt        # Attainment
    ws['B7'].number_format = pct_fmt        # Attrition Rate
    ws['B8'].number_format = int_fmt        # Ramp Months
    ws['B9'].number_format = pct_fmt        # Ramp Productivity

    # Light alternating row fill for readability
    light_fill = PatternFill(start_color='FFD6E4F0', end_color='FFD6E4F0', fill_type='solid')
    for row in range(2, 13):
        if row % 2 == 0:
            for col in range(1, 3):
                ws.cell(row=row, column=col).fill = light_fill

    # Section separator: calculated fields label
    sep_font = Font(name='Calibri', size=10, italic=True, color='808080')
    # Mark the derived rows with a subtle indicator in column A
    for row_idx in [5, 6, 10, 11, 12]:
        ws.cell(row=row_idx, column=1).font = Font(name='Calibri', size=11, italic=True)

    # Column widths
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 20

    # Thin border around data area
    thin_side = Side(style='thin', color='000000')
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    for row in range(1, 13):
        for col in range(1, 3):
            ws.cell(row=row, column=col).border = border

    # --- Additional context sheet: Assumptions ---
    ws2 = wb.create_sheet('Assumptions')
    ws2['A1'] = 'Key Assumptions'
    ws2['A1'].font = Font(name='Calibri', size=14, bold=True)
    assumptions = [
        'Revenue target based on FY2026 board-approved plan',
        'Quota per rep derived from historical average of top 60% performers',
        'Attainment rate reflects trailing 4-quarter weighted average',
        'Attrition rate based on 3-year rolling average for sales org',
        'Ramp time assumes structured onboarding program (6 months to full productivity)',
        'Ramp productivity at 50% means new hires produce half their quota during ramp',
    ]
    for i, text in enumerate(assumptions, 3):
        ws2.cell(row=i, column=1, value=f'{i-2}. {text}')
        ws2.cell(row=i, column=1).font = Font(name='Calibri', size=10)
    ws2.column_dimensions['A'].width = 80

    # --- Additional context sheet: Historical Data ---
    ws3 = wb.create_sheet('Historical')
    ws3['A1'] = 'Year'
    ws3['B1'] = 'Revenue'
    ws3['C1'] = 'Headcount'
    ws3['D1'] = 'Avg Attainment'
    ws3['E1'] = 'Attrition'
    for c in range(1, 6):
        ws3.cell(row=1, column=c).font = Font(bold=True)

    hist_data = [
        [2022, 6500000, 16, 0.82, 0.18],
        [2023, 7200000, 18, 0.80, 0.16],
        [2024, 8500000, 20, 0.84, 0.14],
        [2025, 9200000, 22, 0.86, 0.13],
    ]
    for r, row_data in enumerate(hist_data, 2):
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r, column=c, value=val)
    ws3['B2'].number_format = currency_fmt
    ws3['B3'].number_format = currency_fmt
    ws3['B4'].number_format = currency_fmt
    ws3['B5'].number_format = currency_fmt
    ws3['D2'].number_format = pct_fmt
    ws3['D3'].number_format = pct_fmt
    ws3['D4'].number_format = pct_fmt
    ws3['D5'].number_format = pct_fmt
    ws3['E2'].number_format = pct_fmt
    ws3['E3'].number_format = pct_fmt
    ws3['E4'].number_format = pct_fmt
    ws3['E5'].number_format = pct_fmt
    for c_letter in ['A', 'B', 'C', 'D', 'E']:
        ws3.column_dimensions[c_letter].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
