"""
Reward Script: Sales capacity planning model with formulas
Task ID: calc_sales_072
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20) - B5 Effective Quota formula =B3*B4
  Component 2 (0.20) - B6 Reps Needed formula =ROUNDUP(B2/B5,0)
  Component 3 (0.20) - B10 Replacement Reps formula =ROUNDUP(B6*B7,0)
  Component 4 (0.20) - B11 Ramp Adjustment formula =B10*(1-B9)
  Component 5 (0.20) - B12 Total Headcount formula =B6+ROUNDUP(B11,0)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_072'


def normalize_formula(val):
    """Normalize a formula string for comparison: uppercase, strip spaces."""
    if not isinstance(val, str):
        return None
    return val.upper().replace(" ", "")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that 'Capacity' sheet exists
    if 'Capacity' not in wb.sheetnames:
        print("FAIL: 'Capacity' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Capacity']

    # Component 1: B5 Effective Quota formula =B3*B4 (0.20 points)
    try:
        val = ws['B5'].value
        norm = normalize_formula(val)
        if norm is not None and norm == '=B3*B4':
            print(f"PASS: Component 1 -- B5 has correct formula: {val} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 -- Expected =B3*B4 in B5, found: {repr(val)}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: B6 Reps Needed formula =ROUNDUP(B2/B5,0) (0.20 points)
    try:
        val = ws['B6'].value
        norm = normalize_formula(val)
        if norm is not None and norm == '=ROUNDUP(B2/B5,0)':
            print(f"PASS: Component 2 -- B6 has correct formula: {val} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 -- Expected =ROUNDUP(B2/B5,0) in B6, found: {repr(val)}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: B10 Replacement Reps formula =ROUNDUP(B6*B7,0) (0.20 points)
    try:
        val = ws['B10'].value
        norm = normalize_formula(val)
        if norm is not None and norm == '=ROUNDUP(B6*B7,0)':
            print(f"PASS: Component 3 -- B10 has correct formula: {val} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 3 -- Expected =ROUNDUP(B6*B7,0) in B10, found: {repr(val)}")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: B11 Ramp Adjustment formula =B10*(1-B9) (0.20 points)
    try:
        val = ws['B11'].value
        norm = normalize_formula(val)
        if norm is not None and norm == '=B10*(1-B9)':
            print(f"PASS: Component 4 -- B11 has correct formula: {val} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 4 -- Expected =B10*(1-B9) in B11, found: {repr(val)}")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # Component 5: B12 Total Headcount formula =B6+ROUNDUP(B11,0) (0.20 points)
    try:
        val = ws['B12'].value
        norm = normalize_formula(val)
        if norm is not None and norm == '=B6+ROUNDUP(B11,0)':
            print(f"PASS: Component 5 -- B12 has correct formula: {val} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 5 -- Expected =B6+ROUNDUP(B11,0) in B12, found: {repr(val)}")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
