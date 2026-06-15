"""
Reward Script: Calculate sales cycle velocity for closed deals
Task ID: calc_sales_deal_velocity_065
Domain: libreoffice_calc
Scoring:
  - Component 1: Cycle Days formula (=E-D) in F2:F201               (0.25 pts)
  - Component 2: Size Bucket IFS formula in G2:G201                   (0.25 pts)
  - Component 3: Speed Flag IFS formula in H2:H201                    (0.20 pts)
  - Component 4: Conditional formatting on H column (Fast/Slow/Normal)(0.10 pts)
  - Component 5: VelocityStats formulas (AVERAGEIFS/MINIFS/MAXIFS/MEDIAN) (0.20 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_sales_deal_velocity_065'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: required sheets must exist
    if 'WonDeals' not in wb.sheetnames:
        print("CRITICAL: 'WonDeals' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    if 'VelocityStats' not in wb.sheetnames:
        print("CRITICAL: 'VelocityStats' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws_won = wb['WonDeals']
    ws_vs = wb['VelocityStats']

    # -----------------------------------------------------------------------
    # Component 1: Cycle Days formula in F2:F201 — =E<n>-D<n>  (0.25 points)
    # Task requires: F2:F201 = =E2-D2 style date subtraction formulas
    # Initial state: F2:F201 are all None (empty)
    # -----------------------------------------------------------------------
    try:
        f_formula_count = 0
        f_total = 0
        for row in range(2, 202):
            f_total += 1
            val = ws_won.cell(row=row, column=6).value
            if val is not None and isinstance(val, str):
                # Accept any formula that subtracts D from E (date difference)
                # Pattern: =E<n>-D<n> (case insensitive, spaces stripped)
                normalized = val.strip().upper().replace(' ', '')
                expected = f'=E{row}-D{row}'
                if normalized == expected.upper():
                    f_formula_count += 1
        ratio = f_formula_count / f_total if f_total > 0 else 0.0
        if ratio >= 0.95:
            print(f"PASS: Component 1 — Cycle Days formula F2:F201: {f_formula_count}/{f_total} rows have =E-D formula (0.25 pts)")
            total_score += 0.25
        elif ratio >= 0.5:
            partial = 0.12
            print(f"PARTIAL: Component 1 — Cycle Days formula F2:F201: {f_formula_count}/{f_total} rows have =E-D formula ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Cycle Days formula F2:F201: only {f_formula_count}/{f_total} rows have =E-D formula (expected >=190)")
    except Exception as e:
        print(f"ERROR: Component 1 — Cycle Days formula: {e}")

    # -----------------------------------------------------------------------
    # Component 2: Size Bucket IFS formula in G2:G201 (0.25 points)
    # Task requires: G2:G201 = IFS(C<n><50000,"Small",C<n><200000,"Mid",C<n>>=200000,"Large")
    # Initial state: G2:G201 are all None (empty)
    # -----------------------------------------------------------------------
    try:
        g_formula_count = 0
        g_total = 0
        for row in range(2, 202):
            g_total += 1
            val = ws_won.cell(row=row, column=7).value
            if val is not None and isinstance(val, str):
                normalized = val.strip().upper().replace(' ', '')
                # Must be an IFS formula referencing column C with Small/Mid/Large
                if (normalized.startswith('=IFS(') and
                        f'C{row}<50000' in normalized.replace(' ', '').upper() and
                        '"SMALL"' in normalized and '"MID"' in normalized and '"LARGE"' in normalized):
                    g_formula_count += 1
        ratio = g_formula_count / g_total if g_total > 0 else 0.0
        if ratio >= 0.95:
            print(f"PASS: Component 2 — Size Bucket IFS formula G2:G201: {g_formula_count}/{g_total} rows have correct IFS formula (0.25 pts)")
            total_score += 0.25
        elif ratio >= 0.5:
            partial = 0.12
            print(f"PARTIAL: Component 2 — Size Bucket IFS formula G2:G201: {g_formula_count}/{g_total} rows have IFS formula ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Size Bucket IFS formula G2:G201: only {g_formula_count}/{g_total} rows (expected >=190)")
    except Exception as e:
        print(f"ERROR: Component 2 — Size Bucket formula: {e}")

    # -----------------------------------------------------------------------
    # Component 3: Speed Flag IFS formula in H2:H201 (0.20 points)
    # Task requires: H2:H201 = IFS(F<n><14,"Fast",F<n>>180,"Slow",F<n>>=14,"Normal")
    # Initial state: H2:H201 are all None (empty)
    # -----------------------------------------------------------------------
    try:
        h_formula_count = 0
        h_total = 0
        for row in range(2, 202):
            h_total += 1
            val = ws_won.cell(row=row, column=8).value
            if val is not None and isinstance(val, str):
                normalized = val.strip().upper().replace(' ', '')
                # Must be an IFS formula referencing column F with Fast/Slow/Normal
                if (normalized.startswith('=IFS(') and
                        '"FAST"' in normalized and '"SLOW"' in normalized and
                        '"NORMAL"' in normalized and f'F{row}' in normalized.replace(' ', '').upper()):
                    h_formula_count += 1
        ratio = h_formula_count / h_total if h_total > 0 else 0.0
        if ratio >= 0.95:
            print(f"PASS: Component 3 — Speed Flag IFS formula H2:H201: {h_formula_count}/{h_total} rows have correct IFS formula (0.20 pts)")
            total_score += 0.20
        elif ratio >= 0.5:
            partial = 0.10
            print(f"PARTIAL: Component 3 — Speed Flag IFS formula H2:H201: {h_formula_count}/{h_total} rows have IFS formula ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Speed Flag IFS formula H2:H201: only {h_formula_count}/{h_total} rows (expected >=190)")
    except Exception as e:
        print(f"ERROR: Component 3 — Speed Flag formula: {e}")

    # -----------------------------------------------------------------------
    # Component 4: Conditional formatting on H2:H201 for Fast/Slow/Normal (0.10 points)
    # Task requires: green for Fast, red for Slow, gray/light for Normal
    # Initial state: no conditional formatting on column H
    # -----------------------------------------------------------------------
    try:
        cf_rules = ws_won.conditional_formatting
        # Count how many of the 3 expected label rules are present (Fast, Slow, Normal)
        fast_count = 0
        slow_count = 0
        normal_count = 0

        for cf_range in cf_rules:
            range_str = str(cf_range)
            # Check if this conditional formatting applies to column H area
            if 'H' in range_str:
                for rule in cf_range.rules:
                    if hasattr(rule, 'formula') and rule.formula:
                        formula_upper = ' '.join(str(f) for f in rule.formula).upper()
                        if '"FAST"' in formula_upper:
                            fast_count += 1
                        if '"SLOW"' in formula_upper:
                            slow_count += 1
                        if '"NORMAL"' in formula_upper:
                            normal_count += 1

        cf_conditions_met = sum([
            1 if fast_count > 0 else 0,
            1 if slow_count > 0 else 0,
            1 if normal_count > 0 else 0,
        ])
        if cf_conditions_met >= 3:
            print(f"PASS: Component 4 — Conditional formatting on H column: Fast={fast_count}, Slow={slow_count}, Normal={normal_count} rules found (0.10 pts)")
            total_score += 0.10
        elif cf_conditions_met >= 1:
            partial = round(cf_conditions_met / 3 * 0.10, 2)
            print(f"PARTIAL: Component 4 — Conditional formatting partially met: {cf_conditions_met}/3 rules found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — No conditional formatting for Fast/Slow/Normal found on column H")
    except Exception as e:
        print(f"ERROR: Component 4 — Conditional formatting: {e}")

    # -----------------------------------------------------------------------
    # Component 5: VelocityStats formulas (AVERAGEIFS/MINIFS/MAXIFS/MEDIAN) (0.20 points)
    # Task requires: VelocityStats has AVERAGEIFS, MINIFS, MAXIFS, MEDIAN formulas
    # for 8 reps (rows 5-12) and 3 size buckets (rows 16-18) — total 44 formula cells
    # Initial state: VelocityStats B:E columns for rows 5-12 and 16-18 are all None
    # -----------------------------------------------------------------------
    try:
        vs_formula_count = 0
        vs_expected = 0

        # Rows 5-12 (8 reps) with columns B, C, D, E (4 formula columns each)
        rep_rows = list(range(5, 13))  # 8 rows
        bucket_rows = [16, 17, 18]     # 3 rows
        stat_cols = [2, 3, 4, 5]       # B, C, D, E

        for row in rep_rows:
            for col in stat_cols:
                vs_expected += 1
                val = ws_vs.cell(row=row, column=col).value
                if val is not None and isinstance(val, str):
                    val_upper = val.strip().upper()
                    # Check for the required formula types
                    if (val_upper.startswith('=AVERAGEIFS(') or
                            val_upper.startswith('=MINIFS(') or
                            val_upper.startswith('=MAXIFS(') or
                            val_upper.startswith('=MEDIAN(')):
                        vs_formula_count += 1

        for row in bucket_rows:
            for col in stat_cols:
                vs_expected += 1
                val = ws_vs.cell(row=row, column=col).value
                if val is not None and isinstance(val, str):
                    val_upper = val.strip().upper()
                    if (val_upper.startswith('=AVERAGEIFS(') or
                            val_upper.startswith('=MINIFS(') or
                            val_upper.startswith('=MAXIFS(') or
                            val_upper.startswith('=MEDIAN(')):
                        vs_formula_count += 1

        ratio = vs_formula_count / vs_expected if vs_expected > 0 else 0.0
        if ratio >= 0.95:
            print(f"PASS: Component 5 — VelocityStats formulas: {vs_formula_count}/{vs_expected} formula cells present (0.20 pts)")
            total_score += 0.20
        elif ratio >= 0.5:
            partial = round(ratio * 0.20, 2)
            print(f"PARTIAL: Component 5 — VelocityStats formulas: {vs_formula_count}/{vs_expected} formula cells ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — VelocityStats formulas: only {vs_formula_count}/{vs_expected} formula cells present (expected >=41)")
    except Exception as e:
        print(f"ERROR: Component 5 — VelocityStats formulas: {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
