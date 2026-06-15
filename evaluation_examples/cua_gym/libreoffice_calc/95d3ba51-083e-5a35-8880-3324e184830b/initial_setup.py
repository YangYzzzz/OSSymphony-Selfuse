"""
Initial Setup: Sales deal velocity tracking workbook
Task ID: calc_sales_deal_velocity_065
Domain: libreoffice_calc

Creates a workbook with:
- Sheet 'WonDeals': 200 won deals with Deal ID, Rep, Deal Size, Lead Date, Close Date
  Columns F (Cycle Days), G (Size Bucket), H (Speed Flag) are EMPTY
- Sheet 'VelocityStats': structure for velocity stats by rep and size bucket, but with empty formula cells
"""

import os
import random
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_deal_velocity_065'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

random.seed(42)

REPS = [
    'Alice Nguyen', 'Ben Okafor', 'Carmen Lopez', 'David Kim',
    'Elena Petrov', 'Frank Osei', 'Grace Tan', 'Henry Walsh'
]

DEAL_SIZES = {
    'small': (5000, 49999),
    'mid':   (50000, 199999),
    'large': (200000, 750000)
}

def rand_deal_size():
    tier = random.choices(['small', 'mid', 'large'], weights=[5, 3, 2])[0]
    lo, hi = DEAL_SIZES[tier]
    return random.randint(lo, hi)

def rand_dates():
    """Return (lead_date, close_date) ensuring close >= lead + 1."""
    lead = date(2023, 1, 1) + timedelta(days=random.randint(0, 700))
    cycle_days = random.choices(
        list(range(1, 400)),
        weights=[max(1, 200 - abs(d - 45)) for d in range(1, 400)]
    )[0]
    close = lead + timedelta(days=cycle_days)
    return lead, close

def create_initial():
    wb = openpyxl.Workbook()

    # ---- Sheet 1: WonDeals ----
    ws = wb.active
    ws.title = 'WonDeals'

    headers = ['Deal ID', 'Rep', 'Deal Size', 'Lead Date', 'Close Date',
               'Cycle Days', 'Size Bucket', 'Speed Flag']
    header_font = Font(name='Calibri', bold=True, size=11)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font_white = Font(name='Calibri', bold=True, size=11, color='FFFFFFFF')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Column widths
    col_widths = [14, 18, 14, 14, 14, 13, 14, 13]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'

    # Generate 200 won deals
    random.seed(42)
    for row_idx in range(2, 202):
        deal_id = f'DEAL-{1000 + row_idx - 2:04d}'
        rep = random.choice(REPS)
        deal_size = rand_deal_size()
        lead_date, close_date = rand_dates()

        ws.cell(row=row_idx, column=1, value=deal_id)
        ws.cell(row=row_idx, column=2, value=rep)
        ws.cell(row=row_idx, column=3, value=deal_size)
        ws.cell(row=row_idx, column=4, value=lead_date)
        ws.cell(row=row_idx, column=5, value=close_date)
        # Columns F (6), G (7), H (8) intentionally left EMPTY

        # Format date cells
        ws.cell(row=row_idx, column=4).number_format = 'yyyy-mm-dd'
        ws.cell(row=row_idx, column=5).number_format = 'yyyy-mm-dd'
        # Format deal size as currency
        ws.cell(row=row_idx, column=3).number_format = '$#,##0'

    # ---- Sheet 2: VelocityStats ----
    ws2 = wb.create_sheet('VelocityStats')

    # Title
    ws2['A1'] = 'Sales Cycle Velocity Statistics'
    ws2['A1'].font = Font(name='Calibri', bold=True, size=14)

    # Section 1: Stats by Rep
    ws2['A3'] = 'By Sales Rep'
    ws2['A3'].font = Font(name='Calibri', bold=True, size=12)

    rep_headers = ['Rep', 'Avg Cycle Days', 'Min Cycle Days', 'Max Cycle Days', 'Median Approx']
    rep_header_fill = PatternFill(start_color='FF70AD47', end_color='FF70AD47', fill_type='solid')
    rep_header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFFFF')
    for col, h in enumerate(rep_headers, 1):
        cell = ws2.cell(row=4, column=col, value=h)
        cell.font = rep_header_font
        cell.fill = rep_header_fill
        cell.alignment = Alignment(horizontal='center')

    # Rep rows — values left empty (formulas go here during golden patch)
    for i, rep in enumerate(REPS, 5):
        ws2.cell(row=i, column=1, value=rep)
        # Columns B-E: empty (will hold AVERAGEIFS, MINIFS, MAXIFS, MEDIAN formulas)

    # Section 2: Stats by Size Bucket
    ws2['A14'] = 'By Deal Size Bucket'
    ws2['A14'].font = Font(name='Calibri', bold=True, size=12)

    bucket_headers = ['Size Bucket', 'Avg Cycle Days', 'Min Cycle Days', 'Max Cycle Days', 'Median Approx']
    bucket_header_fill = PatternFill(start_color='FFED7D31', end_color='FFED7D31', fill_type='solid')
    bucket_header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFFFF')
    for col, h in enumerate(bucket_headers, 1):
        cell = ws2.cell(row=15, column=col, value=h)
        cell.font = bucket_header_font
        cell.fill = bucket_header_fill
        cell.alignment = Alignment(horizontal='center')

    buckets = ['Small', 'Mid', 'Large']
    for i, bucket in enumerate(buckets, 16):
        ws2.cell(row=i, column=1, value=bucket)
        # Columns B-E: empty

    # Column widths for VelocityStats
    vs_widths = [20, 16, 16, 16, 16]
    for i, w in enumerate(vs_widths, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

create_initial()
