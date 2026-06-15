"""
Initial Setup: Contract Renewal Management Tracker
Task ID: calc_sales_contract_renewal_056
Domain: libreoffice_calc

Creates an initial spreadsheet with:
- 'Contracts' sheet: 150 contracts with Contract ID, Customer, ARR, Start Date, End Date, (empty) Days to Expiry, (empty) Urgency, Rep
- 'RenewalSummary' sheet: urgency tier labels in A column, empty B and C columns
"""

import os
from datetime import date, timedelta
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_contract_renewal_056'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

# Seed for reproducibility
random.seed(42)

# Realistic data pools
FIRST_NAMES = [
    'Sarah', 'Marcus', 'Jennifer', 'David', 'Emily', 'James', 'Lisa', 'Robert',
    'Michelle', 'Kevin', 'Patricia', 'Thomas', 'Angela', 'Christopher', 'Rachel',
    'William', 'Amanda', 'Daniel', 'Stephanie', 'Michael', 'Karen', 'Jason',
    'Laura', 'Brian', 'Rebecca', 'Andrew', 'Nicole', 'Matthew', 'Samantha', 'Ryan'
]
LAST_NAMES = [
    'Chen', 'Johnson', 'Williams', 'Brown', 'Davis', 'Miller', 'Wilson', 'Moore',
    'Taylor', 'Anderson', 'Thomas', 'Jackson', 'White', 'Harris', 'Martin',
    'Thompson', 'Garcia', 'Martinez', 'Robinson', 'Clark', 'Rodriguez', 'Lewis',
    'Lee', 'Walker', 'Hall', 'Allen', 'Young', 'Hernandez', 'King', 'Wright'
]

COMPANY_PREFIXES = [
    'Apex', 'Nexus', 'Vortex', 'Summit', 'Pinnacle', 'Horizon', 'Fusion',
    'Catalyst', 'Meridian', 'Vertex', 'Stellar', 'Quantum', 'Prism', 'Axiom',
    'Momentum', 'Synergy', 'Elevate', 'Streamline', 'Optima', 'Landmark'
]
COMPANY_SUFFIXES = [
    'Solutions', 'Technologies', 'Consulting', 'Dynamics', 'Systems', 'Industries',
    'Enterprises', 'Analytics', 'Partners', 'Services', 'Group', 'Corp', 'Inc',
    'Global', 'International'
]

REPS = [
    'Alex Morgan', 'Jordan Lee', 'Taylor Brooks', 'Casey Stone', 'Riley Evans',
    'Drew Carter', 'Cameron West', 'Blake Turner', 'Avery Mills', 'Quinn Hayes'
]

def generate_company_name(seed_val):
    random.seed(seed_val)
    prefix = random.choice(COMPANY_PREFIXES)
    suffix = random.choice(COMPANY_SUFFIXES)
    return f'{prefix} {suffix}'

def generate_rep(seed_val):
    random.seed(seed_val * 7 + 13)
    return random.choice(REPS)

def generate_arr(seed_val):
    random.seed(seed_val * 3 + 17)
    # ARR between $12,000 and $480,000
    tier = random.choice(['small', 'small', 'mid', 'mid', 'mid', 'large'])
    if tier == 'small':
        return round(random.randint(12, 60) * 1000)
    elif tier == 'mid':
        return round(random.randint(60, 200) * 1000)
    else:
        return round(random.randint(200, 480) * 1000)

def create_initial():
    wb = openpyxl.Workbook()

    # -------------------------------------------------------
    # Sheet 1: Contracts
    # -------------------------------------------------------
    ws_contracts = wb.active
    ws_contracts.title = 'Contracts'

    # Headers in row 1
    headers = ['Contract ID', 'Customer', 'ARR', 'Start Date', 'End Date',
               'Days to Expiry', 'Urgency', 'Rep']
    for col, header in enumerate(headers, 1):
        cell = ws_contracts.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, name='Calibri', size=11)

    # Column widths
    ws_contracts.column_dimensions['A'].width = 22
    ws_contracts.column_dimensions['B'].width = 30
    ws_contracts.column_dimensions['C'].width = 16
    ws_contracts.column_dimensions['D'].width = 14
    ws_contracts.column_dimensions['E'].width = 14
    ws_contracts.column_dimensions['F'].width = 16
    ws_contracts.column_dimensions['G'].width = 12
    ws_contracts.column_dimensions['H'].width = 20

    today = date.today()

    # Generate 150 contract rows
    # Distribute end dates: roughly 50 in 1-90 days, 50 in 91-180 days, 50 in 181-450 days
    end_date_offsets = (
        [random.randint(1, 90) for _ in range(50)] +
        [random.randint(91, 180) for _ in range(50)] +
        [random.randint(181, 450) for _ in range(50)]
    )
    random.seed(100)
    random.shuffle(end_date_offsets)

    for i in range(150):
        row = i + 2
        contract_id = f'CNT-2024-{i + 1001:04d}'
        customer = generate_company_name(i * 11 + 3)
        arr = generate_arr(i * 5 + 7)
        days_offset = end_date_offsets[i]
        end_date = today + timedelta(days=days_offset)
        # Start date: 1-3 years before end date
        random.seed(i * 13 + 29)
        contract_years = random.choice([1, 2, 3])
        start_date = end_date - timedelta(days=contract_years * 365)
        rep = generate_rep(i * 7 + 11)

        ws_contracts.cell(row=row, column=1, value=contract_id)
        ws_contracts.cell(row=row, column=2, value=customer)

        arr_cell = ws_contracts.cell(row=row, column=3, value=arr)
        arr_cell.number_format = '$#,##0'

        start_cell = ws_contracts.cell(row=row, column=4, value=start_date)
        start_cell.number_format = 'yyyy-mm-dd'

        end_cell = ws_contracts.cell(row=row, column=5, value=end_date)
        end_cell.number_format = 'yyyy-mm-dd'

        # Columns F (Days to Expiry) and G (Urgency) intentionally EMPTY
        # F2:F151 and G2:G151 left blank

        ws_contracts.cell(row=row, column=8, value=rep)

    # Freeze header row
    ws_contracts.freeze_panes = 'A2'

    # -------------------------------------------------------
    # Sheet 2: RenewalSummary
    # -------------------------------------------------------
    ws_summary = wb.create_sheet('RenewalSummary')

    # Column headers in row 1
    ws_summary.cell(row=1, column=1, value='Urgency Tier').font = Font(bold=True, name='Calibri', size=11)
    ws_summary.cell(row=1, column=2, value='Total ARR').font = Font(bold=True, name='Calibri', size=11)
    ws_summary.cell(row=1, column=3, value='Count').font = Font(bold=True, name='Calibri', size=11)

    # Tier labels in A column (as specified in context)
    ws_summary.cell(row=2, column=1, value='Critical (0-90 days)')
    ws_summary.cell(row=3, column=1, value='90-180 days')
    ws_summary.cell(row=4, column=1, value='180+ days')

    # B2:C4 intentionally EMPTY — to be filled by agent with SUMIFS and COUNTIFS

    ws_summary.column_dimensions['A'].width = 24
    ws_summary.column_dimensions['B'].width = 18
    ws_summary.column_dimensions['C'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Contracts sheet: 150 rows, columns A-H (F and G empty)')
    print(f'  RenewalSummary sheet: tier labels in A2:A4, B and C columns empty')

create_initial()
