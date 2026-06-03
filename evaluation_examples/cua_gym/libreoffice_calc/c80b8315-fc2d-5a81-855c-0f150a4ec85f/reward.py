"""
Reward Script: Build a contract renewal management tracker
Task ID: calc_sales_contract_renewal_056
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: F2:F151 formulas =EN-TODAY() (0.25 pts)
  Component 2: G2:G151 IFS urgency formulas (0.25 pts)
  Component 3: Conditional formatting on F2:F151 - red<=90, yellow<=180, green>180 (0.25 pts)
  Component 4: RenewalSummary SUMIFS and COUNTIFS formulas in B2:C4 (0.25 pts)
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_contract_renewal_056'


def verify_task(file_path):
    """
    Verify contract renewal tracker completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Required sheets must exist
    if 'Contracts' not in wb.sheetnames:
        print("FAIL: Sheet 'Contracts' not found - cannot proceed")
        print("REWARD: 0.0")
        return 0.0

    if 'RenewalSummary' not in wb.sheetnames:
        print("FAIL: Sheet 'RenewalSummary' not found - cannot proceed")
        print("REWARD: 0.0")
        return 0.0

    ws_contracts = wb['Contracts']
    ws_summary = wb['RenewalSummary']

    # Component 1: F2:F151 contain =EN-TODAY() formulas (0.25 points)
    # These cells should all have the pattern =E{row}-TODAY()
    # In initial file, F2:F151 are all None
    try:
        f_formula_count = 0
        f_formula_pattern_ok = 0
        f_formula_errors = []

        for row in range(2, 152):
            f_val = ws_contracts.cell(row=row, column=6).value
            if f_val is not None:
                f_formula_count += 1
                # Normalize the formula: should match =E{row}-TODAY()
                formula_str = str(f_val).strip().upper().replace(' ', '')
                expected = f'=E{row}-TODAY()'
                if formula_str == expected.upper():
                    f_formula_pattern_ok += 1
                else:
                    if len(f_formula_errors) < 3:
                        f_formula_errors.append(f"Row {row}: found {repr(f_val)}, expected {expected}")

        if f_formula_count == 150 and f_formula_pattern_ok == 150:
            print(f"PASS: Component 1 — All 150 F column formulas are =E{{row}}-TODAY() (0.25 pts)")
            total_score += 0.25
        elif f_formula_count == 150 and f_formula_pattern_ok >= 140:
            # Partial: mostly correct
            print(f"PARTIAL: Component 1 — {f_formula_pattern_ok}/150 F column formulas correct (0.12 pts)")
            print(f"  Sample errors: {f_formula_errors[:3]}")
            total_score += 0.12
        else:
            print(f"FAIL: Component 1 — Expected 150 =E{{row}}-TODAY() formulas in F2:F151")
            print(f"  Found {f_formula_count} non-None values, {f_formula_pattern_ok} with correct pattern")
            if f_formula_errors:
                print(f"  Sample errors: {f_formula_errors[:3]}")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not check F column formulas: {e}")

    # Component 2: G2:G151 contain IFS urgency formulas (0.25 points)
    # Should be: =IFS(F{row}<=90,"Critical",F{row}<=180,"Watch",F{row}>180,"Healthy")
    try:
        g_formula_count = 0
        g_formula_pattern_ok = 0
        g_formula_errors = []

        for row in range(2, 152):
            g_val = ws_contracts.cell(row=row, column=7).value
            if g_val is not None:
                g_formula_count += 1
                formula_str = str(g_val).strip().upper().replace(' ', '')
                # Check for IFS pattern with Critical/Watch/Healthy and correct F references
                # Pattern: =IFS(F{row}<=90,"CRITICAL",F{row}<=180,"WATCH",F{row}>180,"HEALTHY")
                expected_pattern = f'=IFS(F{row}<=90,"CRITICAL",F{row}<=180,"WATCH",F{row}>180,"HEALTHY")'
                if formula_str == expected_pattern:
                    g_formula_pattern_ok += 1
                else:
                    if len(g_formula_errors) < 3:
                        g_formula_errors.append(f"Row {row}: found {repr(g_val)}")

        if g_formula_count == 150 and g_formula_pattern_ok == 150:
            print(f"PASS: Component 2 — All 150 G column IFS urgency formulas correct (0.25 pts)")
            total_score += 0.25
        elif g_formula_count == 150 and g_formula_pattern_ok >= 140:
            print(f"PARTIAL: Component 2 — {g_formula_pattern_ok}/150 G column IFS formulas correct (0.12 pts)")
            print(f"  Sample errors: {g_formula_errors[:3]}")
            total_score += 0.12
        else:
            print(f"FAIL: Component 2 — Expected 150 IFS formulas in G2:G151")
            print(f"  Found {g_formula_count} non-None values, {g_formula_pattern_ok} with correct pattern")
            if g_formula_errors:
                print(f"  Sample errors: {g_formula_errors[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not check G column formulas: {e}")

    # Component 3: Conditional formatting on F2:F151 — red<=90, yellow<=180, green>180 (0.25 points)
    # Expected: 3 cellIs rules with specific colors
    # red (FFFF0000) for <=90, yellow (FFFFFF00) for <=180, green (FF00FF00) for >180
    try:
        cf_rules_found = []
        has_cf_on_f_column = False

        for cf_range in ws_contracts.conditional_formatting:
            cf_str = str(cf_range)
            # Check if the range covers F2:F151
            if 'F2' in cf_str and 'F151' in cf_str:
                has_cf_on_f_column = True
                cf_obj = ws_contracts.conditional_formatting[cf_range]
                for rule in cf_obj:
                    if rule.type == 'cellIs':
                        try:
                            rule_formula = rule.formula[0] if rule.formula else None
                            rule_op = rule.operator
                            fill_color = rule.dxf.fill.fgColor.rgb if rule.dxf and rule.dxf.fill else None
                            cf_rules_found.append({
                                'formula': rule_formula,
                                'operator': rule_op,
                                'color': fill_color
                            })
                        except Exception as e_rule:
                            print(f"  CF rule error: {e_rule}")

        if has_cf_on_f_column and len(cf_rules_found) >= 3:
            # Check for the 3 required rules
            has_red_90 = False    # red for <=90
            has_yellow_180 = False  # yellow for <=180 (lessThanOrEqual)
            has_green_180 = False   # green for >180

            for rule in cf_rules_found:
                color = rule.get('color', '').upper() if rule.get('color') else ''
                op = rule.get('operator', '').lower() if rule.get('operator') else ''
                formula = str(rule.get('formula', '')).strip()

                # Red fill for <=90 days (Critical)
                if 'FF0000' in color and op in ('lessthanorequal',) and formula == '90':
                    has_red_90 = True

                # Yellow fill for <=180 days (Watch)
                if 'FFFF00' in color and op in ('lessthanorequal',) and formula == '180':
                    has_yellow_180 = True

                # Green fill for >180 days (Healthy)
                if '00FF00' in color and op in ('greaterthan',) and formula == '180':
                    has_green_180 = True

            conditions_met = sum([has_red_90, has_yellow_180, has_green_180])

            if conditions_met == 3:
                print(f"PASS: Component 3 — Conditional formatting on F2:F151 with all 3 color rules (0.25 pts)")
                total_score += 0.25
            elif conditions_met >= 1:
                partial = 0.25 * conditions_met / 3
                partial = round(partial * 4) / 4  # round to nearest 0.25
                print(f"PARTIAL: Component 3 — {conditions_met}/3 CF rules correct ({partial} pts)")
                print(f"  red<=90: {has_red_90}, yellow<=180: {has_yellow_180}, green>180: {has_green_180}")
                print(f"  Rules found: {cf_rules_found}")
                total_score += partial
            else:
                print(f"FAIL: Component 3 — CF rules found but none match expected patterns")
                print(f"  Rules found: {cf_rules_found}")
        else:
            print(f"FAIL: Component 3 — No conditional formatting on F2:F151 found")
            print(f"  has_cf_on_f: {has_cf_on_f_column}, rules_found: {len(cf_rules_found)}")
    except Exception as e:
        print(f"ERROR: Component 3 — Could not check conditional formatting: {e}")

    # Component 4: RenewalSummary SUMIFS and COUNTIFS formulas in B2:C4 (0.25 points)
    # B2: =SUMIFS(Contracts!C:C,Contracts!G:G,"Critical")
    # B3: =SUMIFS(Contracts!C:C,Contracts!G:G,"Watch")
    # B4: =SUMIFS(Contracts!C:C,Contracts!G:G,"Healthy")
    # C2: =COUNTIFS(Contracts!G:G,"Critical")
    # C3: =COUNTIFS(Contracts!G:G,"Watch")
    # C4: =COUNTIFS(Contracts!G:G,"Healthy")
    try:
        summary_checks = 0
        summary_total = 6
        summary_errors = []

        # Define expected formulas (normalized)
        expected_formulas = {
            (2, 2): 'SUMIFS',    # B2 has SUMIFS for Critical
            (2, 3): 'COUNTIFS',  # C2 has COUNTIFS for Critical
            (3, 2): 'SUMIFS',    # B3 has SUMIFS for Watch
            (3, 3): 'COUNTIFS',  # C3 has COUNTIFS for Watch
            (4, 2): 'SUMIFS',    # B4 has SUMIFS for Healthy
            (4, 3): 'COUNTIFS',  # C4 has COUNTIFS for Healthy
        }
        urgency_map = {2: 'CRITICAL', 3: 'WATCH', 4: 'HEALTHY'}

        for (row, col), func_type in expected_formulas.items():
            cell_val = ws_summary.cell(row=row, column=col).value
            if cell_val is None:
                summary_errors.append(f"  Cell ({row},{col}) is None, expected {func_type} formula")
                continue

            formula_upper = str(cell_val).upper().replace(' ', '')
            urgency = urgency_map.get(row, '')

            # Check that it uses the right function and references the right urgency tier
            if func_type in formula_upper and urgency in formula_upper and 'CONTRACTS' in formula_upper:
                summary_checks += 1
            else:
                summary_errors.append(f"  Cell ({row},{col}): found {repr(cell_val)}, expected {func_type} referencing {urgency}")

        if summary_checks == summary_total:
            print(f"PASS: Component 4 — All 6 SUMIFS/COUNTIFS formulas in RenewalSummary (0.25 pts)")
            total_score += 0.25
        elif summary_checks >= 4:
            print(f"PARTIAL: Component 4 — {summary_checks}/{summary_total} RenewalSummary formulas correct (0.12 pts)")
            for err in summary_errors:
                print(err)
            total_score += 0.12
        else:
            print(f"FAIL: Component 4 — Only {summary_checks}/{summary_total} RenewalSummary formulas correct")
            for err in summary_errors:
                print(err)
    except Exception as e:
        print(f"ERROR: Component 4 — Could not check RenewalSummary formulas: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
