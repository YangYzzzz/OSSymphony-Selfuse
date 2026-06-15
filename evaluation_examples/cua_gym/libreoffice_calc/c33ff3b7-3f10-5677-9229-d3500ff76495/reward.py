"""
Reward Script: Build a pivot table summarizing total budget allocation per project phase.
Task ID: calc_pivot_013
Domain: libreoffice_calc
Scoring:
  Component 1 (0.15): PivotTable sheet exists
  Component 2 (0.25): Phase labels present and correct
  Component 3 (0.40): Amount values correct per phase
  Component 4 (0.20): Grand Total row present and correct
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_013'

# Expected phase totals from task context
EXPECTED_PHASES = {
    'Planning': 45000,
    'Development': 180000,
    'Testing': 95000,
    'Deployment': 60000,
}
EXPECTED_GRAND_TOTAL = 380000


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: PivotTable sheet exists (0.15 points)
    # This is a task-introduced change: initial has only 'Budget', golden adds 'PivotTable'
    try:
        pivot_sheet = None
        for name in wb.sheetnames:
            if 'pivot' in name.lower():
                pivot_sheet = wb[name]
                break
        if pivot_sheet is not None:
            print(f"PASS: Component 1 — PivotTable sheet found: '{pivot_sheet.title}' (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — No pivot table sheet found. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0  # No pivot sheet means nothing else to check
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print("REWARD: 0.0")
        return 0.0

    # Read all data from pivot sheet into a dict: phase_name -> amount
    # We search flexibly for phase labels and their associated amounts
    ws = pivot_sheet
    found_phases = {}
    grand_total_value = None

    try:
        for r in range(1, ws.max_row + 1):
            cell_a = ws.cell(row=r, column=1).value
            if cell_a is None:
                continue
            cell_a_str = str(cell_a).strip()

            # Check for grand total row
            if 'grand total' in cell_a_str.lower() or 'total' == cell_a_str.lower():
                # Look for the numeric value in columns 2 onward
                for c in range(2, ws.max_column + 1):
                    val = ws.cell(row=r, column=c).value
                    if isinstance(val, (int, float)):
                        grand_total_value = val
                        break
                continue

            # Check if this is a phase label
            for phase_name in EXPECTED_PHASES:
                if cell_a_str.lower() == phase_name.lower():
                    # Find the associated amount in the same row
                    for c in range(2, ws.max_column + 1):
                        val = ws.cell(row=r, column=c).value
                        if isinstance(val, (int, float)):
                            found_phases[phase_name] = val
                            break
                    break
    except Exception as e:
        print(f"ERROR: Reading pivot sheet data — {e}")

    # Component 2: Phase labels present and correct (0.25 points)
    # Award partial credit: 0.0625 per phase found (4 phases)
    try:
        phases_found_count = 0
        for phase_name in EXPECTED_PHASES:
            if phase_name in found_phases:
                phases_found_count += 1
            else:
                print(f"FAIL: Component 2 — Phase '{phase_name}' not found in pivot table")

        if phases_found_count == 4:
            print(f"PASS: Component 2 — All 4 phase labels found (0.25 pts)")
            total_score += 0.25
        elif phases_found_count > 0:
            phase_label_score = (phases_found_count / 4.0) * 0.25
            print(f"PARTIAL: Component 2 — {phases_found_count}/4 phase labels found ({phase_label_score:.4f} pts)")
            total_score += phase_label_score
        else:
            print(f"FAIL: Component 2 — No phase labels found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Amount values correct per phase (0.40 points)
    # Award 0.10 per correct phase amount
    try:
        correct_amounts = 0
        for phase_name, expected_val in EXPECTED_PHASES.items():
            if phase_name in found_phases:
                actual_val = found_phases[phase_name]
                if abs(float(actual_val) - float(expected_val)) < 1.0:
                    correct_amounts += 1
                    print(f"PASS: Component 3 — {phase_name} amount = {actual_val} (expected {expected_val})")
                else:
                    print(f"FAIL: Component 3 — {phase_name} amount = {actual_val}, expected {expected_val}")
            else:
                print(f"FAIL: Component 3 — {phase_name} not found, cannot check amount")

        if correct_amounts == 4:
            print(f"PASS: Component 3 — All 4 phase amounts correct (0.40 pts)")
            total_score += 0.40
        elif correct_amounts > 0:
            amount_score = (correct_amounts / 4.0) * 0.40
            print(f"PARTIAL: Component 3 — {correct_amounts}/4 amounts correct ({amount_score:.4f} pts)")
            total_score += amount_score
        else:
            print(f"FAIL: Component 3 — No amounts correct")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Grand Total present and correct (0.20 points)
    try:
        if grand_total_value is not None:
            if abs(float(grand_total_value) - EXPECTED_GRAND_TOTAL) < 1.0:
                print(f"PASS: Component 4 — Grand Total = {grand_total_value} (expected {EXPECTED_GRAND_TOTAL}) (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 4 — Grand Total = {grand_total_value}, expected {EXPECTED_GRAND_TOTAL}")
        else:
            print(f"FAIL: Component 4 — Grand Total row not found in pivot table")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
