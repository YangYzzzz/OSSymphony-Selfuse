"""
Initial Setup: Build a pivot table summarizing total budget allocation per project phase.
Task ID: calc_pivot_013
Domain: libreoffice_calc

Creates a Budget sheet with 80 rows of project budget data across 3 projects
and 4 phases. Amounts are carefully distributed to match ground truth totals.
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_013'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Budget'

    # --- Headers ---
    headers = ['LineItem', 'Project', 'Phase', 'Category', 'Amount']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Data Design ---
    # Target totals: Planning=45000, Development=180000, Testing=95000, Deployment=60000
    # Total=380000, 80 rows
    # Distribution: ~15 Planning, ~30 Development, ~20 Testing, ~15 Deployment = 80

    projects = ['Alpha', 'Beta', 'Gamma']

    categories_by_phase = {
        'Planning': ['Requirements Analysis', 'Stakeholder Review', 'Scope Definition',
                      'Risk Assessment', 'Resource Planning', 'Timeline Development'],
        'Development': ['Backend Engineering', 'Frontend Development', 'API Integration',
                        'Database Design', 'Code Review', 'Architecture Design',
                        'DevOps Setup', 'Middleware Development'],
        'Testing': ['Unit Testing', 'Integration Testing', 'UAT', 'Performance Testing',
                    'Security Audit', 'Regression Testing'],
        'Deployment': ['Server Provisioning', 'CI/CD Pipeline', 'Documentation',
                       'Training Materials', 'Go-Live Support', 'Monitoring Setup'],
    }

    # We'll generate exact amounts that sum to the target for each phase
    phase_targets = {
        'Planning': (15, 45000),
        'Development': (30, 180000),
        'Testing': (20, 95000),
        'Deployment': (15, 60000),
    }

    rows = []
    for phase, (count, target_total) in phase_targets.items():
        # Generate 'count' amounts that sum to target_total
        # Use a simple approach: generate count-1 random amounts, last one is remainder
        amounts = []
        avg = target_total / count
        remaining = target_total
        for i in range(count - 1):
            # Random amount between 40% and 160% of average, rounded to nearest 50
            low = max(100, int(avg * 0.4))
            high = int(avg * 1.6)
            amt = random.randint(low, high)
            amt = round(amt / 50) * 50  # round to nearest 50
            # Ensure we don't overshoot
            if remaining - amt < (count - 1 - i) * 100:
                amt = 100
            amounts.append(amt)
            remaining -= amt

        # Last amount is the remainder
        amounts.append(remaining)

        cats = categories_by_phase[phase]
        for i in range(count):
            proj = projects[i % len(projects)]
            cat = cats[i % len(cats)]
            rows.append((proj, phase, cat, amounts[i]))

    # Shuffle rows to make them look natural (but keep deterministic)
    random.shuffle(rows)

    # Write data
    for r, (proj, phase, cat, amt) in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=r - 1)  # LineItem number
        ws.cell(row=r, column=2, value=proj)
        ws.cell(row=r, column=3, value=phase)
        ws.cell(row=r, column=4, value=cat)
        ws.cell(row=r, column=5, value=amt)

    # Format Amount column as currency
    for r in range(2, 82):
        ws.cell(row=r, column=5).number_format = '$#,##0'

    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Verify totals
    phase_sums = {}
    for r in range(2, 82):
        phase = ws.cell(row=r, column=3).value
        amt = ws.cell(row=r, column=5).value
        phase_sums[phase] = phase_sums.get(phase, 0) + amt
    print(f'Phase totals: {phase_sums}')
    print(f'Grand total: {sum(phase_sums.values())}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
