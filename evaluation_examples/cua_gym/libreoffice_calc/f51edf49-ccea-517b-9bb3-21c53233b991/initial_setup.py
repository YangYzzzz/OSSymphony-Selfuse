"""
Initial Setup: Navigate to December sheet using Navigator panel
Task ID: calc_gsi_074
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_074'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # Define 20 sheet names - months plus additional business sheets
    sheet_names = [
        'January', 'February', 'March', 'April', 'May', 'June',
        'July', 'August', 'September', 'October', 'November', 'December',
        'Q1 Summary', 'Q2 Summary', 'Q3 Summary', 'Q4 Summary',
        'Annual Overview', 'Budget Forecast', 'Headcount', 'KPI Dashboard'
    ]

    # Header style
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    # Monthly data templates
    departments = ['Engineering', 'Marketing', 'Sales', 'Finance', 'HR',
                   'Operations', 'Legal', 'Customer Support', 'Product', 'Design']
    employees = [
        ('Sarah Chen', 'Engineering', 92500),
        ('Marcus Johnson', 'Marketing', 78200),
        ('Priya Patel', 'Sales', 85600),
        ('David Kim', 'Finance', 88900),
        ('Elena Rodriguez', 'HR', 73400),
        ('James Wilson', 'Operations', 81200),
        ('Aisha Mohammed', 'Legal', 95800),
        ('Ryan O\'Brien', 'Customer Support', 67500),
        ('Lisa Wang', 'Product', 91300),
        ('Tom Nakamura', 'Design', 79800),
        ('Rachel Foster', 'Engineering', 87600),
        ('Carlos Mendez', 'Sales', 82100),
    ]

    # Create the first sheet (rename default)
    ws = wb.active
    ws.title = sheet_names[0]

    # Create remaining sheets
    for name in sheet_names[1:]:
        wb.create_sheet(name)

    # Populate monthly sheets (January through December)
    for month_idx in range(12):
        ws = wb[sheet_names[month_idx]]
        headers = ['Employee', 'Department', 'Base Salary', 'Bonus', 'Total Comp', 'Hours Logged', 'Projects']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Set column widths
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 12

        # Add employee data with month-based variation
        for r, (name, dept, base) in enumerate(employees, 2):
            # Vary bonus by month index for realism
            bonus = round(base * (0.05 + 0.01 * month_idx) * (0.8 + 0.4 * ((r + month_idx) % 5) / 4), 2)
            total = base + bonus
            hours = 140 + ((r * 7 + month_idx * 13) % 40)
            projects = 1 + ((r + month_idx) % 4)

            ws.cell(row=r, column=1, value=name).border = thin_border
            ws.cell(row=r, column=2, value=dept).border = thin_border
            c = ws.cell(row=r, column=3, value=base)
            c.number_format = '$#,##0.00'
            c.border = thin_border
            c = ws.cell(row=r, column=4, value=bonus)
            c.number_format = '$#,##0.00'
            c.border = thin_border
            c = ws.cell(row=r, column=5, value=total)
            c.number_format = '$#,##0.00'
            c.border = thin_border
            ws.cell(row=r, column=6, value=hours).border = thin_border
            ws.cell(row=r, column=7, value=projects).border = thin_border

    # Populate quarterly summary sheets (Q1-Q4)
    for q_idx in range(4):
        ws = wb[f'Q{q_idx+1} Summary']
        q_headers = ['Department', 'Total Salary', 'Total Bonus', 'Avg Hours', 'Headcount']
        for col, h in enumerate(q_headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12

        for r, dept in enumerate(departments, 2):
            ws.cell(row=r, column=1, value=dept)
            ws.cell(row=r, column=2, value=round(250000 + (r * 31417 + q_idx * 7919) % 150000, 2))
            ws.cell(row=r, column=3, value=round(15000 + (r * 2713 + q_idx * 1997) % 25000, 2))
            ws.cell(row=r, column=4, value=round(155 + (r * 3 + q_idx * 7) % 20, 1))
            ws.cell(row=r, column=5, value=3 + (r + q_idx) % 8)

    # Annual Overview
    ws = wb['Annual Overview']
    ao_headers = ['Metric', 'Q1', 'Q2', 'Q3', 'Q4', 'Annual Total']
    for col, h in enumerate(ao_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws.column_dimensions['A'].width = 22
    metrics = ['Total Revenue', 'Operating Expenses', 'Net Profit', 'Employee Count',
               'Customer Satisfaction', 'Churn Rate', 'New Customers', 'Support Tickets']
    base_vals = [1250000, 890000, 360000, 48, 4.2, 0.032, 340, 1250]
    for r, (metric, base) in enumerate(zip(metrics, base_vals), 2):
        ws.cell(row=r, column=1, value=metric)
        for q in range(4):
            variation = 1 + 0.05 * q + 0.02 * r
            ws.cell(row=r, column=q+2, value=round(base * variation, 2))
        ws.cell(row=r, column=6, value=round(base * 4.3, 2))

    # Budget Forecast
    ws = wb['Budget Forecast']
    bf_headers = ['Category', '2025 Actual', '2026 Projected', 'Variance', 'Notes']
    for col, h in enumerate(bf_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['E'].width = 30
    categories = [
        ('Salaries', 4200000, 4536000, 'Annual 8% increase'),
        ('Benefits', 1260000, 1360800, 'Healthcare cost up 5%'),
        ('Office Rent', 480000, 504000, 'Lease renewal in Q3'),
        ('Software Licenses', 180000, 210000, 'New DevOps tools'),
        ('Travel', 95000, 120000, 'Conference budget expanded'),
        ('Marketing', 350000, 420000, 'Brand campaign launch'),
        ('Equipment', 220000, 195000, 'Reduced refresh cycle'),
        ('Training', 85000, 110000, 'Leadership program'),
        ('Insurance', 145000, 152000, 'Standard adjustment'),
        ('Miscellaneous', 60000, 72000, 'Buffer increase'),
    ]
    for r, (cat, actual, projected, note) in enumerate(categories, 2):
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=2, value=actual).number_format = '$#,##0'
        ws.cell(row=r, column=3, value=projected).number_format = '$#,##0'
        ws.cell(row=r, column=4, value=projected - actual).number_format = '$#,##0'
        ws.cell(row=r, column=5, value=note)

    # Headcount
    ws = wb['Headcount']
    hc_headers = ['Department', 'Current', 'Open Positions', 'Target', 'Fill Rate']
    for col, h in enumerate(hc_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws.column_dimensions['A'].width = 20
    for r, dept in enumerate(departments, 2):
        current = 5 + (r * 3) % 15
        openpos = 1 + r % 3
        target = current + openpos
        ws.cell(row=r, column=1, value=dept)
        ws.cell(row=r, column=2, value=current)
        ws.cell(row=r, column=3, value=openpos)
        ws.cell(row=r, column=4, value=target)
        ws.cell(row=r, column=5, value=round(current / target, 2)).number_format = '0.00%'

    # KPI Dashboard
    ws = wb['KPI Dashboard']
    kpi_headers = ['KPI', 'Target', 'Actual', 'Status', 'Trend']
    for col, h in enumerate(kpi_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    kpis = [
        ('Revenue Growth', '15%', '12.8%', 'Below Target', 'Improving'),
        ('Customer NPS', '70', '73', 'On Track', 'Stable'),
        ('Employee Retention', '92%', '89%', 'At Risk', 'Declining'),
        ('Sprint Velocity', '45 pts', '48 pts', 'Exceeding', 'Improving'),
        ('Bug Resolution Time', '24h', '18h', 'Exceeding', 'Improving'),
        ('Uptime SLA', '99.9%', '99.95%', 'On Track', 'Stable'),
        ('Cost per Acquisition', '$45', '$52', 'Below Target', 'Worsening'),
        ('MRR Growth', '8%', '9.2%', 'Exceeding', 'Improving'),
    ]
    for r, (kpi, target, actual, status, trend) in enumerate(kpis, 2):
        ws.cell(row=r, column=1, value=kpi)
        ws.cell(row=r, column=2, value=target)
        ws.cell(row=r, column=3, value=actual)
        ws.cell(row=r, column=4, value=status)
        ws.cell(row=r, column=5, value=trend)

    # Set January as the active sheet (first sheet)
    wb.active = 0

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the workbook in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
