"""
Reward Script: Revenue waterfall analysis with formulas
Task ID: calc_sales_059
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): C2 contains formula =B2 (Running Total start)
  Component 2 (0.30): C3:C6 contain running total formulas (=prev+current)
  Component 3 (0.20): B7 contains formula =C6 (Ending ARR)
  Component 4 (0.25): D3:D6 contain % of Starting ARR formulas (=Bx/B$2)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_059'


def normalize_formula(f):
    """Normalize a formula string for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Waterfall' sheet must exist
    if 'Waterfall' not in wb.sheetnames:
        print("FAIL: 'Waterfall' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Waterfall']

    # Component 1: C2 contains formula =B2 (Running Total start) (0.25 points)
    try:
        c2_val = ws['C2'].value
        if isinstance(c2_val, str) and normalize_formula(c2_val) == '=B2':
            print(f"PASS: Component 1 -- C2 has formula =B2 (found: {c2_val}) (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 -- Expected formula =B2 in C2, found: {repr(c2_val)}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: C3:C6 contain running total formulas (0.30 points)
    # Expected: C3=C2+B3, C4=C3+B4, C5=C4+B5, C6=C5+B6
    try:
        expected_running = {
            'C3': '=C2+B3',
            'C4': '=C3+B4',
            'C5': '=C4+B5',
            'C6': '=C5+B6',
        }
        running_pass = 0
        for cell_ref, expected_formula in expected_running.items():
            actual = ws[cell_ref].value
            if isinstance(actual, str) and normalize_formula(actual) == normalize_formula(expected_formula):
                running_pass += 1
                print(f"  PASS: {cell_ref} has correct formula {actual}")
            else:
                print(f"  FAIL: {cell_ref} expected {expected_formula}, found: {repr(actual)}")

        if running_pass == 4:
            print(f"PASS: Component 2 -- All 4 running total formulas correct (0.30 pts)")
            total_score += 0.30
        elif running_pass > 0:
            partial = round(0.30 * running_pass / 4, 2)
            print(f"PARTIAL: Component 2 -- {running_pass}/4 running total formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- No running total formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: B7 contains formula =C6 (Ending ARR) (0.20 points)
    try:
        b7_val = ws['B7'].value
        if isinstance(b7_val, str) and normalize_formula(b7_val) == '=C6':
            print(f"PASS: Component 3 -- B7 has formula =C6 (found: {b7_val}) (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 3 -- Expected formula =C6 in B7, found: {repr(b7_val)}")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: D3:D6 contain % of Starting ARR formulas (0.25 points)
    # Expected: D3=B3/B$2, D4=B4/B$2, D5=B5/B$2, D6=B6/B$2
    try:
        expected_pct = {
            'D3': '=B3/B$2',
            'D4': '=B4/B$2',
            'D5': '=B5/B$2',
            'D6': '=B6/B$2',
        }
        pct_pass = 0
        for cell_ref, expected_formula in expected_pct.items():
            actual = ws[cell_ref].value
            if isinstance(actual, str) and normalize_formula(actual) == normalize_formula(expected_formula):
                pct_pass += 1
                print(f"  PASS: {cell_ref} has correct formula {actual}")
            else:
                print(f"  FAIL: {cell_ref} expected {expected_formula}, found: {repr(actual)}")

        if pct_pass == 4:
            print(f"PASS: Component 4 -- All 4 percentage formulas correct (0.25 pts)")
            total_score += 0.25
        elif pct_pass > 0:
            partial = round(0.25 * pct_pass / 4, 2)
            print(f"PARTIAL: Component 4 -- {pct_pass}/4 percentage formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 -- No percentage formulas found")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved GUI state before verification
def persist_app_state(domain):
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
