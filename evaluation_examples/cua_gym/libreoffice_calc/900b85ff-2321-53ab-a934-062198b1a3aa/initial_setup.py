"""
Initial Setup: Revenue waterfall analysis spreadsheet with raw data only.
Task ID: calc_sales_059
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_059'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Waterfall ---
    ws = wb.active
    ws.title = 'Waterfall'

    # Headers
    headers = ['Category', 'Amount', 'Running Total', '% of Starting ARR']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    white_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows (raw data only - NO formulas)
    data = [
        ['Starting ARR', 2000000],
        ['New Business', 450000],
        ['Expansion', 280000],
        ['Contraction', -120000],
        ['Churn', -310000],
        ['Ending ARR', None],  # B7 must be empty - task asks agent to add formula
    ]

    for r, (category, amount) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=category)
        if amount is not None:
            cell_b = ws.cell(row=r, column=2, value=amount)
            cell_b.number_format = '$#,##0'

    # Column C (Running Total) and D (% of Starting ARR) left empty intentionally
    # The task asks the agent to build these formulas

    # Set column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 22

    # Add a second sheet with supporting context data to make the workbook realistic
    ws2 = wb.create_sheet('Monthly Detail')
    detail_headers = ['Month', 'New Business', 'Expansion', 'Contraction', 'Churn', 'Net Change']
    for col, h in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', size=11, bold=True)

    monthly_data = [
        ['Jan 2025', 38000, 22000, -9500, -25000, 25500],
        ['Feb 2025', 35000, 24000, -10000, -28000, 21000],
        ['Mar 2025', 42000, 26000, -11000, -24000, 33000],
        ['Apr 2025', 37000, 21000, -8500, -27000, 22500],
        ['May 2025', 40000, 25000, -10500, -26000, 28500],
        ['Jun 2025', 36000, 23000, -9000, -25500, 24500],
        ['Jul 2025', 39000, 24500, -10000, -26500, 27000],
        ['Aug 2025', 41000, 22500, -11500, -27500, 24500],
        ['Sep 2025', 35000, 23500, -10000, -25000, 23500],
        ['Oct 2025', 37000, 24000, -10500, -26000, 24500],
        ['Nov 2025', 34000, 23000, -9500, -24500, 23000],
        ['Dec 2025', 36000, 22500, -10000, -25000, 23500],
    ]

    for r, row_data in enumerate(monthly_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if c >= 2:
                cell.number_format = '$#,##0'

    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws2.column_dimensions[col_letter].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
