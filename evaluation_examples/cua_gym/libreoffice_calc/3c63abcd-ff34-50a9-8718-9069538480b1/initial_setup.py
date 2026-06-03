"""
Initial Setup: Add a hyperlink in cell B5 on the 'Index' sheet that jumps to Data.A1
Task ID: calc_gg1_033
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_033'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Index (navigation hub) ---
    ws_index = wb.active
    ws_index.title = 'Index'

    # Title row
    ws_index.merge_cells('A1:D1')
    ws_index['A1'] = 'Project Navigation Hub'
    ws_index['A1'].font = Font(name='Arial', size=16, bold=True, color='1F4E79')
    ws_index['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Subtitle
    ws_index.merge_cells('A2:D2')
    ws_index['A2'] = 'Q1 2025 Operations Dashboard'
    ws_index['A2'].font = Font(name='Arial', size=11, italic=True, color='4472C4')
    ws_index['A2'].alignment = Alignment(horizontal='center')

    # Section header
    ws_index['A4'] = 'Section'
    ws_index['B4'] = 'Quick Link'
    ws_index['C4'] = 'Description'
    ws_index['D4'] = 'Last Updated'
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    for col_letter in ['A', 'B', 'C', 'D']:
        cell = ws_index[f'{col_letter}4']
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Navigation rows
    ws_index['A5'] = 'Data Repository'
    # B5 is intentionally LEFT EMPTY - this is where the agent must add the hyperlink
    ws_index['C5'] = 'Contains all project metrics and KPIs'
    ws_index['D5'] = '2025-03-28'

    ws_index['A6'] = 'Team Directory'
    ws_index['B6'] = '(Coming Soon)'
    ws_index['C6'] = 'Contact info for all project members'
    ws_index['D6'] = '2025-03-15'

    ws_index['A7'] = 'Budget Overview'
    ws_index['B7'] = '(Coming Soon)'
    ws_index['C7'] = 'Financial allocations and expenditure tracking'
    ws_index['D7'] = '2025-03-20'

    ws_index['A8'] = 'Timeline & Milestones'
    ws_index['B8'] = '(Coming Soon)'
    ws_index['C8'] = 'Project deadlines and deliverable dates'
    ws_index['D8'] = '2025-03-22'

    # Footer note
    ws_index['A10'] = 'Note: Quick links will be added as sheets become available.'
    ws_index['A10'].font = Font(name='Arial', size=9, italic=True, color='808080')

    # Column widths
    ws_index.column_dimensions['A'].width = 22
    ws_index.column_dimensions['B'].width = 22
    ws_index.column_dimensions['C'].width = 45
    ws_index.column_dimensions['D'].width = 16

    # --- Sheet 2: Data ---
    ws_data = wb.create_sheet('Data')

    # Headers
    data_headers = ['Employee', 'Department', 'Region', 'Q1 Revenue', 'Q1 Expenses',
                    'Net Profit', 'Projects Completed', 'Satisfaction Score']
    header_font_data = Font(name='Arial', size=11, bold=True)
    header_fill_data = PatternFill(start_color='FFD9E2F3', end_color='FFD9E2F3', fill_type='solid')
    for col, h in enumerate(data_headers, 1):
        cell = ws_data.cell(row=1, column=col, value=h)
        cell.font = header_font_data
        cell.fill = header_fill_data
        cell.alignment = Alignment(horizontal='center')

    # Realistic data
    data_rows = [
        ['Sarah Chen', 'Engineering', 'West', 125400, 78200, 47200, 8, 4.7],
        ['Marcus Johnson', 'Marketing', 'East', 98700, 62300, 36400, 5, 4.2],
        ['Priya Patel', 'Sales', 'South', 187500, 95600, 91900, 12, 4.8],
        ['David Kim', 'Engineering', 'West', 112300, 71800, 40500, 7, 4.5],
        ['Elena Rodriguez', 'Operations', 'Central', 76800, 54200, 22600, 4, 3.9],
        ['James Wright', 'Sales', 'East', 165200, 88900, 76300, 10, 4.6],
        ['Aisha Mohammed', 'Marketing', 'South', 89400, 58700, 30700, 6, 4.1],
        ['Ryan O\'Brien', 'Engineering', 'Central', 134600, 82400, 52200, 9, 4.4],
        ['Lisa Chang', 'Operations', 'West', 71200, 49800, 21400, 3, 4.0],
        ['Michael Torres', 'Sales', 'East', 152800, 91200, 61600, 11, 4.7],
        ['Hannah Weber', 'Marketing', 'Central', 94300, 61500, 32800, 5, 4.3],
        ['Tomoko Nakamura', 'Engineering', 'South', 118900, 75300, 43600, 8, 4.6],
    ]

    for r, row_data in enumerate(data_rows, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_data.cell(row=r, column=c, value=val)
            # Format currency columns
            if c in [4, 5, 6]:
                cell.number_format = '$#,##0'

    # Column widths for Data sheet
    ws_data.column_dimensions['A'].width = 20
    ws_data.column_dimensions['B'].width = 14
    ws_data.column_dimensions['C'].width = 10
    ws_data.column_dimensions['D'].width = 14
    ws_data.column_dimensions['E'].width = 14
    ws_data.column_dimensions['F'].width = 14
    ws_data.column_dimensions['G'].width = 20
    ws_data.column_dimensions['H'].width = 18

    # Freeze header row
    ws_data.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
