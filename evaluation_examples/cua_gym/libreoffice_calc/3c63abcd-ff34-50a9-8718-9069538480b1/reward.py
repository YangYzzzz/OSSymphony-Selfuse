"""
Reward Script: Add hyperlink in B5 on 'Index' sheet to Data!A1 with display text 'Go to Data Sheet'
Task ID: calc_gg1_033
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): B5 on Index sheet has a hyperlink object
  Component 2 (0.30): B5 display text is 'Go to Data Sheet'
  Component 3 (0.25): Hyperlink location targets Data sheet cell A1
  Component 4 (0.15): Font styling indicates hyperlink (underline + blue-ish color)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_033'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice document."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: workbook must have 'Index' and 'Data' sheets
    if 'Index' not in wb.sheetnames:
        print("CRITICAL: 'Index' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0
    if 'Data' not in wb.sheetnames:
        print("CRITICAL: 'Data' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws_index = wb['Index']
    cell_b5 = ws_index['B5']

    # Component 1: B5 on Index sheet has a hyperlink object (0.30 points)
    try:
        hyperlink = cell_b5.hyperlink
        if hyperlink is not None:
            print(f"PASS: Component 1 -- B5 has a hyperlink object (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 -- B5 has no hyperlink (hyperlink is None)")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: B5 display text is 'Go to Data Sheet' (0.30 points)
    try:
        cell_value = cell_b5.value
        if cell_value is not None and str(cell_value).strip() == 'Go to Data Sheet':
            print(f"PASS: Component 2 -- B5 value is 'Go to Data Sheet' (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 2 -- B5 value is {cell_value!r}, expected 'Go to Data Sheet'")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Hyperlink location targets Data sheet cell A1 (0.25 points)
    try:
        hyperlink = cell_b5.hyperlink
        if hyperlink is not None:
            loc = hyperlink.location or ''
            target = hyperlink.target or ''
            # The hyperlink should reference the 'Data' sheet, cell A1
            # openpyxl stores internal links in 'location' field like "'Data'!A1" or "Data!A1"
            # It could also be in target for external-style references
            combined = (loc + ' ' + target).lower().replace("'", "").replace('"', '')
            has_data_sheet = 'data' in combined
            has_a1 = 'a1' in combined
            if has_data_sheet and has_a1:
                print(f"PASS: Component 3 -- Hyperlink points to Data!A1 (location={loc!r}, target={target!r}) (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 3 -- Hyperlink location={loc!r}, target={target!r}; expected reference to Data!A1")
        else:
            print(f"FAIL: Component 3 -- No hyperlink object to check location")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Font styling indicates hyperlink (underline + blue-ish color) (0.15 points)
    try:
        font = cell_b5.font
        has_underline = font.underline is not None and font.underline != 'none'
        # Check for blue-ish font color (typical hyperlink colors)
        has_blue = False
        try:
            rgb = font.color.rgb
            if rgb is not None:
                # Extract RGB components from ARGB hex string
                rgb_str = str(rgb)
                if len(rgb_str) >= 6:
                    # Take last 6 chars as RGB
                    hex_rgb = rgb_str[-6:]
                    r_val = int(hex_rgb[0:2], 16)
                    g_val = int(hex_rgb[2:4], 16)
                    b_val = int(hex_rgb[4:6], 16)
                    # Blue-ish: blue channel significantly higher than red, or typical hyperlink colors
                    if b_val > r_val and b_val > 100:
                        has_blue = True
        except Exception:
            pass

        if has_underline and has_blue:
            print(f"PASS: Component 4 -- Font has underline={font.underline} and blue color (0.15 pts)")
            total_score += 0.15
        elif has_underline:
            # Partial: underline present but color may differ
            print(f"PARTIAL: Component 4 -- Font has underline but color check uncertain (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 4 -- underline={font.underline}, blue={has_blue}")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
