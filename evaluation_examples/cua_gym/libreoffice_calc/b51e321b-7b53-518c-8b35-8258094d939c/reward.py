"""
Reward Script: Wrap VLOOKUP in IFERROR to show 'Product Not Found'
Task ID: calc_lf_019
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): E2 contains an IFERROR formula
  Component 2 (0.3): The IFERROR wraps a VLOOKUP on D2 against A2:B4
  Component 3 (0.3): The fallback value is "Product Not Found"
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_019'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice document."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Inventory' sheet must exist
    if 'Inventory' not in wb.sheetnames:
        print("FAIL: 'Inventory' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Inventory']
    formula_raw = ws['E2'].value

    # Component 1: E2 contains an IFERROR formula (0.4 points)
    try:
        if formula_raw is not None and isinstance(formula_raw, str):
            formula_norm = formula_raw.upper().replace(" ", "")
            if formula_norm.startswith("=IFERROR("):
                print(f"PASS: Component 1 — E2 contains IFERROR formula (0.4 pts)")
                total_score += 0.4
            else:
                print(f"FAIL: Component 1 — E2 does not start with =IFERROR(, found: {formula_raw}")
        else:
            print(f"FAIL: Component 1 — E2 is empty or not a formula, found: {repr(formula_raw)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: IFERROR wraps a VLOOKUP referencing D2 and A2:B4 (0.3 points)
    try:
        if formula_raw is not None and isinstance(formula_raw, str):
            formula_norm = formula_raw.upper().replace(" ", "")
            # Check that VLOOKUP is present inside IFERROR with the right lookup params
            has_vlookup = "VLOOKUP(" in formula_norm
            has_d2_ref = "D2" in formula_norm
            # Accept various range formats: A2:B4, A:B, $A$2:$B$4, etc.
            has_range = bool(re.search(r'A\$?2:\$?B\$?4|A:B|\$A\$2:\$B\$4', formula_norm))
            has_col_index = "2," in formula_norm  # column index 2
            if has_vlookup and has_d2_ref and has_range and has_col_index:
                print(f"PASS: Component 2 — VLOOKUP(D2,A2:B4,2,...) found (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — VLOOKUP structure mismatch. "
                      f"vlookup={has_vlookup}, d2={has_d2_ref}, range={has_range}, col2={has_col_index}. "
                      f"Formula: {formula_raw}")
        else:
            print(f"FAIL: Component 2 — E2 is not a formula string")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Fallback value is "Product Not Found" (0.3 points)
    try:
        if formula_raw is not None and isinstance(formula_raw, str):
            # Look for the string "Product Not Found" (case-insensitive) as the IFERROR fallback
            # It should appear as the second argument to IFERROR, quoted
            formula_upper = formula_raw.upper()
            if "PRODUCT NOT FOUND" in formula_upper:
                print(f"PASS: Component 3 — Fallback 'Product Not Found' present (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 3 — 'Product Not Found' not found in formula: {formula_raw}")
        else:
            print(f"FAIL: Component 3 — E2 is not a formula string")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
