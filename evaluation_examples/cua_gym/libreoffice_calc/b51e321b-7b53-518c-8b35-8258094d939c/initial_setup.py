"""
Initial Setup: VLOOKUP with IFERROR for product lookup
Task ID: calc_lf_019
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_019'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Inventory ---
    ws = wb.active
    ws.title = 'Inventory'

    # Headers
    ws['A1'] = 'Product'
    ws['B1'] = 'Stock'
    ws['D1'] = 'Search'
    ws['E1'] = 'Stock'

    # Style headers
    header_font = Font(bold=True)
    for cell_ref in ['A1', 'B1', 'D1', 'E1']:
        ws[cell_ref].font = header_font

    # Product data (A2:B4)
    ws['A2'] = 'Pen'
    ws['B2'] = 500
    ws['A3'] = 'Pencil'
    ws['B3'] = 300
    ws['A4'] = 'Eraser'
    ws['B4'] = 200

    # Search value
    ws['D2'] = 'Marker'

    # E2 is intentionally left EMPTY - that's what the task asks the agent to fill

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 4
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
