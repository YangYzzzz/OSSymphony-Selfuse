"""
Initial Setup: Add minor gridlines on the Y-axis to an existing line chart
Task ID: calc_chart_gridlines_023
Domain: libreoffice_calc

Creates a spreadsheet with humidity readings for 8 locations and a line chart
with major Y-axis gridlines but NO minor gridlines (pre-task state).
"""

import os
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import ChartLines

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_gridlines_023'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Humidity ---
    ws = wb.active
    ws.title = 'Humidity'

    # Headers
    ws['A1'] = 'Location'
    ws['B1'] = 'Humidity %'

    # Data rows (8 stations as specified in context)
    data = [
        ('Station A', 68),
        ('Station B', 72),
        ('Station C', 65),
        ('Station D', 79),
        ('Station E', 61),
        ('Station F', 74),
        ('Station G', 83),
        ('Station H', 70),
    ]
    for r, (loc, hum) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=loc)
        ws.cell(row=r, column=2, value=hum)

    # --- Create Line Chart with major gridlines on Y-axis (NO minor gridlines) ---
    chart = LineChart()
    chart.title = 'Humidity by Location'
    chart.y_axis.title = 'Humidity %'
    chart.x_axis.title = 'Location'
    chart.style = 10

    # Data reference (B1:B9 includes header row for title)
    data_ref = Reference(ws, min_col=2, min_row=1, max_row=9)
    chart.add_data(data_ref, titles_from_data=True)

    # Category reference
    cats = Reference(ws, min_col=1, min_row=2, max_row=9)
    chart.set_categories(cats)

    # Add major gridlines on Y-axis (explicitly set)
    chart.y_axis.majorGridlines = ChartLines()

    # NO minor gridlines on Y-axis (pre-task state: minorGridlines = None)
    chart.y_axis.minorGridlines = None

    # Place chart on sheet
    ws.add_chart(chart, 'D2')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
