"""
Initial Setup: Apply conditional formatting to highlight deals over $100,000
Task ID: calc_sales_015
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_015'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: BigDeals ---
    ws = wb.active
    ws.title = 'BigDeals'

    # Headers
    ws['A1'] = 'Deal Name'
    ws['B1'] = 'Value'

    # Data rows - realistic deal names and values
    data = [
        ['Alpha', 45000],
        ['Beta', 120000],
        ['Gamma', 95000],
        ['Delta', 210000],
        ['Epsilon', 88000],
        ['Zeta', 150000],
        ['Eta', 67000],
    ]
    for r, (name, value) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=value)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15

    # NO conditional formatting in initial state
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
