"""
Reward Script: Build a pipeline funnel summary
Task ID: calc_sales_pipeline_funnel_004
Domain: libreoffice_calc
Scoring:
  Component 1: COUNTIFS formulas in FunnelSummary B2:B7 (0.35 pts)
  Component 2: SUMIFS formulas in FunnelSummary C2:C7 (0.30 pts)
  Component 3: A bar/funnel chart present on FunnelSummary (0.35 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — reward scripts always run on the VM
TASK_ID = 'calc_sales_pipeline_funnel_004'

EXPECTED_STAGES = ['Lead', 'Qualified', 'Demo', 'Proposal', 'Negotiation', 'Closed Won']


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook — precondition gate
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: FunnelSummary sheet must exist
    if 'FunnelSummary' not in wb.sheetnames:
        print("CRITICAL: 'FunnelSummary' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['FunnelSummary']

    # Component 1: COUNTIFS formulas in B2:B7 (0.35 points)
    # The task requires COUNTIFS formulas to count deals per stage.
    # Initial file has None in B2:B7; golden file has COUNTIFS formulas.
    try:
        countifs_count = 0
        for row in range(2, 8):
            cell_val = ws.cell(row=row, column=2).value
            if cell_val is not None:
                formula_str = str(cell_val).upper().replace(' ', '')
                if 'COUNTIFS' in formula_str:
                    countifs_count += 1

        if countifs_count == 6:
            print(f"PASS: Component 1 — All 6 COUNTIFS formulas found in B2:B7 (0.35 pts)")
            total_score += 0.35
        elif countifs_count > 0:
            partial = round(0.35 * countifs_count / 6, 4)
            print(f"PARTIAL: Component 1 — {countifs_count}/6 COUNTIFS formulas found in B2:B7 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No COUNTIFS formulas found in B2:B7 (expected 6)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: SUMIFS formulas in C2:C7 (0.30 points)
    # The task requires SUMIFS formulas to total deal values per stage.
    # Initial file has None in C2:C7; golden file has SUMIFS formulas.
    try:
        sumifs_count = 0
        for row in range(2, 8):
            cell_val = ws.cell(row=row, column=3).value
            if cell_val is not None:
                formula_str = str(cell_val).upper().replace(' ', '')
                if 'SUMIFS' in formula_str:
                    sumifs_count += 1

        if sumifs_count == 6:
            print(f"PASS: Component 2 — All 6 SUMIFS formulas found in C2:C7 (0.30 pts)")
            total_score += 0.30
        elif sumifs_count > 0:
            partial = round(0.30 * sumifs_count / 6, 4)
            print(f"PARTIAL: Component 2 — {sumifs_count}/6 SUMIFS formulas found in C2:C7 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No SUMIFS formulas found in C2:C7 (expected 6)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: A bar or funnel chart present on FunnelSummary (0.35 points)
    # The task requires a chart visualizing deal count by stage.
    # Initial file has 0 charts; golden file has 1 BarChart.
    try:
        charts = ws._charts
        bar_or_funnel_chart = False
        chart_details = []
        for chart in charts:
            chart_type_name = type(chart).__name__
            chart_details.append(chart_type_name)
            # Accept BarChart or any chart type (funnel charts may show as BarChart in openpyxl)
            if 'Bar' in chart_type_name or 'bar' in chart_type_name.lower():
                bar_or_funnel_chart = True

        if bar_or_funnel_chart:
            print(f"PASS: Component 3 — Bar/funnel chart found on FunnelSummary: {chart_details} (0.35 pts)")
            total_score += 0.35
        elif len(charts) > 0:
            # Some chart type found — still award partial credit (0.20) for having any chart
            print(f"PARTIAL: Component 3 — Chart found but not bar type: {chart_details} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 3 — No chart found on FunnelSummary (expected a bar/funnel chart)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
