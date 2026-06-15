"""
Initial Setup: Build a pipeline funnel summary
Task ID: calc_sales_pipeline_funnel_004
Domain: libreoffice_calc

Creates a spreadsheet with:
- Sheet 'Deals': 150 deals with ID, Stage, Value, Rep, Region columns
- Sheet 'FunnelSummary': Stage names in A2:A7, headers in B1:C1, empty B2:B7 and C2:C7
  (no formulas yet, no chart — those are the task)
"""

import openpyxl
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_pipeline_funnel_004'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

random.seed(42)

# Sales stages in funnel order
STAGES = ['Lead', 'Qualified', 'Demo', 'Proposal', 'Negotiation', 'Closed Won']

# Realistic distribution: more at top of funnel, fewer at bottom
STAGE_COUNTS = {
    'Lead': 45,
    'Qualified': 35,
    'Demo': 28,
    'Proposal': 22,
    'Negotiation': 12,
    'Closed Won': 8,
}

REPS = [
    'Sarah Chen', 'Marcus Johnson', 'Emily Rodriguez', 'David Kim',
    'Jessica Patel', 'Ryan Thompson', 'Amanda Torres', 'Chris Nguyen',
    'Lauren Williams', 'Michael Brown', 'Stephanie Davis', 'Kevin Martinez',
]

REGIONS = ['North', 'South', 'East', 'West', 'Central']

# Value ranges per stage (higher stage = higher value deals)
VALUE_RANGES = {
    'Lead':        (5000, 25000),
    'Qualified':   (15000, 60000),
    'Demo':        (20000, 90000),
    'Proposal':    (30000, 150000),
    'Negotiation': (50000, 200000),
    'Closed Won':  (40000, 180000),
}


def create_initial():
    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ #
    # Sheet 1: Deals
    # ------------------------------------------------------------------ #
    ws_deals = wb.active
    ws_deals.title = 'Deals'

    # Headers
    headers = ['ID', 'Stage', 'Value', 'Rep', 'Region']
    for col, h in enumerate(headers, 1):
        ws_deals.cell(row=1, column=col, value=h)

    # Generate 150 deals
    all_deals = []
    deal_id = 1001
    for stage, count in STAGE_COUNTS.items():
        lo, hi = VALUE_RANGES[stage]
        for _ in range(count):
            value = round(random.uniform(lo, hi), 2)
            rep = random.choice(REPS)
            region = random.choice(REGIONS)
            all_deals.append((deal_id, stage, value, rep, region))
            deal_id += 1

    # Shuffle so stages are interleaved (more realistic)
    random.shuffle(all_deals)

    for r, (did, stage, value, rep, region) in enumerate(all_deals, 2):
        ws_deals.cell(row=r, column=1, value=did)
        ws_deals.cell(row=r, column=2, value=stage)
        ws_deals.cell(row=r, column=3, value=value)
        ws_deals.cell(row=r, column=4, value=rep)
        ws_deals.cell(row=r, column=5, value=region)

    # Column widths for readability
    ws_deals.column_dimensions['A'].width = 10
    ws_deals.column_dimensions['B'].width = 14
    ws_deals.column_dimensions['C'].width = 14
    ws_deals.column_dimensions['D'].width = 20
    ws_deals.column_dimensions['E'].width = 12

    # ------------------------------------------------------------------ #
    # Sheet 2: FunnelSummary
    # ------------------------------------------------------------------ #
    ws_funnel = wb.create_sheet('FunnelSummary')

    # Header row (row 1)
    ws_funnel.cell(row=1, column=1, value='Stage')
    ws_funnel.cell(row=1, column=2, value='Count')
    ws_funnel.cell(row=1, column=3, value='Total Value')

    # Stage names in A2:A7 (funnel order: Lead → Closed Won)
    for r, stage in enumerate(STAGES, 2):
        ws_funnel.cell(row=r, column=1, value=stage)

    # B2:B7 and C2:C7 are intentionally LEFT EMPTY
    # (The agent's task is to add COUNTIFS and SUMIFS formulas here)

    ws_funnel.column_dimensions['A'].width = 16
    ws_funnel.column_dimensions['B'].width = 10
    ws_funnel.column_dimensions['C'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Deals sheet: 150 rows (rows 2-151), stages: {STAGES}')
    print(f'  FunnelSummary: stages in A2:A7, headers B1/C1, B2:B7 and C2:C7 empty (no formulas, no chart)')


create_initial()
