"""
Initial Setup: Inventory master list with duplicate SKUs for duplicate detection task
Task ID: calc_ops_inventory_duplicate_sku_005
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_inventory_duplicate_sku_005'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'ItemMaster'

    # Headers
    headers = ['SKU', 'Product Name', 'Category', 'Unit of Measure', 'Standard Cost', 'Status']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    # Column G is intentionally empty (no header, no data)

    # 150 rows of inventory data
    # We define 137 unique SKUs and 13 duplicates (second occurrences scattered)
    # Duplicates at row positions (1-indexed data rows): 15, 27, 39, 52, 64, 76, 88, 100, 112, 120, 130, 140, 148
    # These will repeat SKUs from earlier rows

    # Full data list: each entry is (SKU, Product Name, Category, Unit of Measure, Standard Cost, Status)
    data = [
        ('SKU-10001', 'Industrial Safety Gloves', 'Safety Equipment', 'Pair', 12.50, 'Active'),
        ('SKU-10002', 'Heavy-Duty Extension Cord 25ft', 'Electrical', 'Each', 28.99, 'Active'),
        ('SKU-10003', 'Stainless Steel Mixing Bowl Set', 'Kitchen Equipment', 'Set', 45.75, 'Active'),
        ('SKU-10004', 'Wireless Barcode Scanner', 'Technology', 'Each', 189.00, 'Active'),
        ('SKU-10005', 'Adjustable Shelving Unit 72in', 'Storage', 'Each', 210.50, 'Active'),
        ('SKU-10006', 'Disposable Nitrile Gloves Box', 'Safety Equipment', 'Box', 19.99, 'Active'),
        ('SKU-10007', 'PVC Pipe 1-inch x 10ft', 'Plumbing', 'Each', 8.45, 'Active'),
        ('SKU-10008', 'Steel Toe Work Boots Size 10', 'Safety Equipment', 'Pair', 89.95, 'Active'),
        ('SKU-10009', 'Industrial Degreaser 1 Gallon', 'Cleaning Supplies', 'Gallon', 24.30, 'Active'),
        ('SKU-10010', 'Cordless Power Drill 18V', 'Power Tools', 'Each', 145.00, 'Active'),
        ('SKU-10011', 'Packing Tape 3-inch x 110yd', 'Packaging', 'Roll', 5.75, 'Active'),
        ('SKU-10012', 'LED Warehouse Light 150W', 'Lighting', 'Each', 67.50, 'Active'),
        ('SKU-10013', 'Forklift Battery Charger 24V', 'Material Handling', 'Each', 520.00, 'Active'),
        ('SKU-10014', 'Safety Data Sheet Binder', 'Office Supplies', 'Each', 7.25, 'Active'),
        ('SKU-10003', 'Stainless Steel Mixing Bowl Set', 'Kitchen Equipment', 'Set', 45.75, 'Active'),  # DUPLICATE row 15
        ('SKU-10015', 'Conveyor Belt Cleaner Brush', 'Maintenance', 'Each', 38.00, 'Active'),
        ('SKU-10016', 'Hydraulic Hand Pallet Jack', 'Material Handling', 'Each', 385.00, 'Active'),
        ('SKU-10017', 'Anti-Slip Stair Nosing Tape', 'Safety Equipment', 'Roll', 14.50, 'Active'),
        ('SKU-10018', 'Welding Helmet Auto-Darkening', 'Safety Equipment', 'Each', 134.75, 'Active'),
        ('SKU-10019', 'Heavy Duty Zip Ties 100pk', 'Fasteners', 'Pack', 9.99, 'Active'),
        ('SKU-10020', 'Air Compressor 6-Gallon Tank', 'Pneumatic Tools', 'Each', 289.00, 'Active'),
        ('SKU-10021', 'Corrugated Cardboard Boxes 12x12', 'Packaging', 'Case', 42.00, 'Active'),
        ('SKU-10022', 'Label Printer Thermal 4x6', 'Technology', 'Each', 175.00, 'Active'),
        ('SKU-10023', 'Aluminum Step Ladder 8ft', 'Safety Equipment', 'Each', 112.50, 'Active'),
        ('SKU-10024', 'Mop Bucket with Wringer', 'Cleaning Supplies', 'Each', 55.00, 'Active'),
        ('SKU-10025', 'Network Ethernet Cable Cat6 100ft', 'Technology', 'Each', 22.99, 'Active'),
        ('SKU-10008', 'Steel Toe Work Boots Size 10', 'Safety Equipment', 'Pair', 89.95, 'Active'),  # DUPLICATE row 27
        ('SKU-10026', 'Industrial Floor Scale 2000lb', 'Measurement', 'Each', 895.00, 'Active'),
        ('SKU-10027', 'Spray Paint Can Gray Primer', 'Maintenance', 'Can', 6.50, 'Active'),
        ('SKU-10028', 'Bubble Wrap Roll 12in x 100ft', 'Packaging', 'Roll', 31.50, 'Active'),
        ('SKU-10029', 'High-Visibility Safety Vest', 'Safety Equipment', 'Each', 8.75, 'Active'),
        ('SKU-10030', 'Digital Caliper 6-inch', 'Measurement', 'Each', 42.00, 'Active'),
        ('SKU-10031', 'Antifreeze Coolant 1 Gallon', 'Automotive', 'Gallon', 15.99, 'Active'),
        ('SKU-10032', 'Plastic Drum 55 Gallon Blue', 'Storage', 'Each', 78.50, 'Active'),
        ('SKU-10033', 'Angle Grinder 4.5-inch 7A', 'Power Tools', 'Each', 88.00, 'Active'),
        ('SKU-10034', 'Weatherproof Electrical Box', 'Electrical', 'Each', 11.25, 'Active'),
        ('SKU-10035', 'Dust Collector 1HP Shop Vac', 'Power Tools', 'Each', 199.00, 'Active'),
        ('SKU-10036', 'Stretch Wrap Film 18in x 1500ft', 'Packaging', 'Roll', 27.50, 'Active'),
        ('SKU-10016', 'Hydraulic Hand Pallet Jack', 'Material Handling', 'Each', 385.00, 'Active'),  # DUPLICATE row 39
        ('SKU-10037', 'Pipe Wrench 18-inch', 'Hand Tools', 'Each', 34.75, 'Active'),
        ('SKU-10038', 'Industrial Fan 30-inch Belt Drive', 'HVAC', 'Each', 445.00, 'Active'),
        ('SKU-10039', 'Safety Lock Padlock 3-Pack', 'Safety Equipment', 'Pack', 18.50, 'Active'),
        ('SKU-10040', 'Grease Cartridge Multi-Purpose', 'Maintenance', 'Each', 7.99, 'Active'),
        ('SKU-10041', 'Plastic Bin Small 6x4x4', 'Storage', 'Each', 3.25, 'Active'),
        ('SKU-10042', 'Socket Set 40-Piece SAE/Metric', 'Hand Tools', 'Set', 78.00, 'Active'),
        ('SKU-10043', 'Reflective Tape 2-inch x 30ft', 'Safety Equipment', 'Roll', 12.00, 'Active'),
        ('SKU-10044', 'Water Pump Submersible 1/2HP', 'Plumbing', 'Each', 167.50, 'Active'),
        ('SKU-10045', 'Marking Paint Inverted White', 'Maintenance', 'Can', 5.99, 'Active'),
        ('SKU-10046', 'UPS Battery Backup 1500VA', 'Technology', 'Each', 210.00, 'Active'),
        ('SKU-10047', 'Loading Dock Bumper Rubber', 'Material Handling', 'Each', 65.00, 'Active'),
        ('SKU-10048', 'Circuit Breaker 20A Single Pole', 'Electrical', 'Each', 9.50, 'Active'),
        ('SKU-10020', 'Air Compressor 6-Gallon Tank', 'Pneumatic Tools', 'Each', 289.00, 'Active'),  # DUPLICATE row 52
        ('SKU-10049', 'Eye Wash Station Wall-Mount', 'Safety Equipment', 'Each', 145.00, 'Active'),
        ('SKU-10050', 'Forklift Forks 48-inch Class II', 'Material Handling', 'Pair', 320.00, 'Active'),
        ('SKU-10051', 'Paint Roller Cover 9-inch Nap', 'Maintenance', 'Each', 4.50, 'Active'),
        ('SKU-10052', 'Steel Storage Cabinet Lockable', 'Storage', 'Each', 425.00, 'Active'),
        ('SKU-10053', 'Brake Cleaner Spray 14oz', 'Automotive', 'Can', 8.25, 'Active'),
        ('SKU-10054', 'Box Cutter Utility Knife', 'Hand Tools', 'Each', 6.99, 'Active'),
        ('SKU-10055', 'Fiberglass Extension Ladder 24ft', 'Safety Equipment', 'Each', 289.50, 'Active'),
        ('SKU-10056', 'Electric Tape Black 3/4in x 66ft', 'Electrical', 'Roll', 3.75, 'Active'),
        ('SKU-10057', 'Rubber Floor Mat Anti-Fatigue', 'Safety Equipment', 'Each', 35.00, 'Active'),
        ('SKU-10058', 'Concrete Anchor Bolt 3/8in 100pk', 'Fasteners', 'Box', 22.50, 'Active'),
        ('SKU-10059', 'Paint Sprayer Airless 1/4HP', 'Maintenance', 'Each', 199.00, 'Inactive'),
        ('SKU-10060', 'Ratchet Strap 2in x 27ft 4-Pack', 'Material Handling', 'Pack', 34.99, 'Active'),
        ('SKU-10022', 'Label Printer Thermal 4x6', 'Technology', 'Each', 175.00, 'Active'),  # DUPLICATE row 64
        ('SKU-10061', 'Safety Cone 28-inch Orange', 'Safety Equipment', 'Each', 14.75, 'Active'),
        ('SKU-10062', 'Grinding Wheel 4.5-inch Metal', 'Abrasives', 'Each', 5.50, 'Active'),
        ('SKU-10063', 'Industrial Sweeper Push 24-inch', 'Cleaning Supplies', 'Each', 68.00, 'Active'),
        ('SKU-10064', 'Drum Pump Electric 55 Gallon', 'Fluid Handling', 'Each', 245.00, 'Active'),
        ('SKU-10065', 'Torque Wrench 1/2in Drive 25-250ft', 'Hand Tools', 'Each', 98.00, 'Active'),
        ('SKU-10066', 'Hearing Protection Earmuff', 'Safety Equipment', 'Each', 24.50, 'Active'),
        ('SKU-10067', 'Magnetic Tool Tray 11-inch', 'Hand Tools', 'Each', 16.00, 'Active'),
        ('SKU-10068', 'Generator Portable 3500W Gas', 'Power Equipment', 'Each', 680.00, 'Active'),
        ('SKU-10069', 'Work Bench Steel 72x30 Adj Height', 'Furniture', 'Each', 795.00, 'Active'),
        ('SKU-10070', 'HVAC Filter 20x20x2 MERV11', 'HVAC', 'Each', 18.75, 'Active'),
        ('SKU-10071', 'Pallet Wrap Dispenser Handheld', 'Packaging', 'Each', 28.00, 'Active'),
        ('SKU-10072', 'Cutting Disc 4.5-inch Thin', 'Abrasives', 'Each', 2.99, 'Active'),
        ('SKU-10073', 'Oil Filter Removal Tool Set', 'Automotive', 'Set', 22.00, 'Active'),
        ('SKU-10074', 'Fire Extinguisher ABC 5lb', 'Safety Equipment', 'Each', 55.00, 'Active'),
        ('SKU-10075', 'Network Switch 8-Port Gigabit', 'Technology', 'Each', 44.99, 'Active'),
        ('SKU-10034', 'Weatherproof Electrical Box', 'Electrical', 'Each', 11.25, 'Active'),  # DUPLICATE row 76
        ('SKU-10076', 'Stretch Film Machine Grade 80ga', 'Packaging', 'Roll', 38.00, 'Active'),
        ('SKU-10077', 'Caulking Gun Heavy Duty 29oz', 'Hand Tools', 'Each', 19.50, 'Active'),
        ('SKU-10078', 'Spill Kit Oil Only 14-Gallon', 'Environmental', 'Kit', 75.00, 'Active'),
        ('SKU-10079', 'Steel Drum Dolly 55 Gallon', 'Material Handling', 'Each', 95.00, 'Active'),
        ('SKU-10080', 'Conduit Pipe EMT 1/2in x 10ft', 'Electrical', 'Each', 7.85, 'Active'),
        ('SKU-10081', 'Inspection Camera Flexible 9mm', 'Technology', 'Each', 129.00, 'Active'),
        ('SKU-10082', 'Concrete Mix 80lb Bag Fast-Set', 'Construction', 'Bag', 8.99, 'Active'),
        ('SKU-10083', 'Plastic Pallet 48x40 Nestable', 'Material Handling', 'Each', 42.00, 'Active'),
        ('SKU-10084', 'Bench Vise 4-inch Swivel', 'Hand Tools', 'Each', 85.00, 'Active'),
        ('SKU-10085', 'Compressed Air Duster 12oz Can', 'Technology', 'Can', 8.25, 'Active'),
        ('SKU-10086', 'Tarp Heavy Duty 10x12 Blue', 'Storage', 'Each', 16.50, 'Active'),
        ('SKU-10087', 'Lockout Tagout Station 6-Lock', 'Safety Equipment', 'Each', 185.00, 'Active'),
        ('SKU-10043', 'Reflective Tape 2-inch x 30ft', 'Safety Equipment', 'Roll', 12.00, 'Active'),  # DUPLICATE row 88
        ('SKU-10088', 'Chain Hoist 1-Ton Manual', 'Material Handling', 'Each', 145.00, 'Active'),
        ('SKU-10089', 'Solvent Cleaner Acetone 1 Gallon', 'Chemicals', 'Gallon', 19.50, 'Inactive'),
        ('SKU-10090', 'Masking Tape 2-inch x 60yd', 'Packaging', 'Roll', 4.25, 'Active'),
        ('SKU-10091', 'Valve Gate 2-inch Bronze', 'Plumbing', 'Each', 32.50, 'Active'),
        ('SKU-10092', 'Cable Tie Mount Adhesive 100pk', 'Fasteners', 'Pack', 7.75, 'Active'),
        ('SKU-10093', 'Drum Heater Band 55 Gallon', 'Fluid Handling', 'Each', 225.00, 'Active'),
        ('SKU-10094', 'Socket Wrench Set 3/8in Drive', 'Hand Tools', 'Set', 54.00, 'Active'),
        ('SKU-10095', 'RFID Asset Tag Pack of 50', 'Technology', 'Pack', 95.00, 'Active'),
        ('SKU-10096', 'Scaffold Board Steel 7ft', 'Construction', 'Each', 48.50, 'Active'),
        ('SKU-10097', 'Sealant Silicone Clear 10.3oz', 'Construction', 'Tube', 6.99, 'Active'),
        ('SKU-10052', 'Steel Storage Cabinet Lockable', 'Storage', 'Each', 425.00, 'Active'),  # DUPLICATE row 100
        ('SKU-10098', 'Parking Barrier Portable Yellow', 'Safety Equipment', 'Each', 29.50, 'Active'),
        ('SKU-10099', 'Thread Tap Set 40-Piece Metric', 'Hand Tools', 'Set', 45.00, 'Active'),
        ('SKU-10100', 'Welding Wire ER70S-6 2lb Spool', 'Welding', 'Spool', 22.00, 'Active'),
        ('SKU-10101', 'Mig Welding Nozzle 5pk', 'Welding', 'Pack', 14.50, 'Active'),
        ('SKU-10102', 'Power Strip Surge Protector 6-Out', 'Electrical', 'Each', 24.99, 'Active'),
        ('SKU-10103', 'Drum Funnel Stainless Steel', 'Fluid Handling', 'Each', 38.50, 'Active'),
        ('SKU-10104', 'First Aid Kit 100-Piece Wall', 'Safety Equipment', 'Each', 65.00, 'Active'),
        ('SKU-10105', 'Metal Detector Handheld Wand', 'Security', 'Each', 89.00, 'Active'),
        ('SKU-10106', 'Pallet Truck Scale 5500lb Cap', 'Material Handling', 'Each', 1250.00, 'Active'),
        ('SKU-10107', 'Drill Bit Set 29-Piece HSS', 'Power Tools', 'Set', 38.00, 'Active'),
        ('SKU-10060', 'Ratchet Strap 2in x 27ft 4-Pack', 'Material Handling', 'Pack', 34.99, 'Active'),  # DUPLICATE row 112
        ('SKU-10108', 'Epoxy Floor Coating Gray 1 Gal', 'Construction', 'Gallon', 55.00, 'Active'),
        ('SKU-10109', 'Respirator N95 Particulate 20pk', 'Safety Equipment', 'Pack', 24.99, 'Active'),
        ('SKU-10110', 'Pipe Insulation Foam 3/4in x 6ft', 'Plumbing', 'Each', 3.50, 'Active'),
        ('SKU-10111', 'Band Saw Blade 93.5in 6TPI', 'Power Tools', 'Each', 28.50, 'Active'),
        ('SKU-10112', 'Packing Foam Sheet 2in x 24x72', 'Packaging', 'Each', 18.75, 'Active'),
        ('SKU-10113', 'Tow Chain Grade 70 20ft x 5/16in', 'Material Handling', 'Each', 75.00, 'Active'),
        ('SKU-10114', 'Wire Mesh Basket 18x12x9', 'Storage', 'Each', 22.00, 'Active'),
        ('SKU-10115', 'Battery 9V Alkaline 12pk', 'Electrical', 'Pack', 12.99, 'Active'),
        ('SKU-10078', 'Spill Kit Oil Only 14-Gallon', 'Environmental', 'Kit', 75.00, 'Active'),  # DUPLICATE row 120
        ('SKU-10116', 'Padlock Combination 1-inch Brass', 'Security', 'Each', 11.50, 'Active'),
        ('SKU-10117', 'Drill Chuck Key 3/8in', 'Power Tools', 'Each', 5.99, 'Active'),
        ('SKU-10118', 'Mop Head Cotton 24oz Industrial', 'Cleaning Supplies', 'Each', 9.75, 'Active'),
        ('SKU-10119', 'Cylinder Valve Cap Plastic 10pk', 'Fluid Handling', 'Pack', 6.25, 'Active'),
        ('SKU-10120', 'Label Holder Clear Adhesive 2in', 'Office Supplies', 'Pack', 8.50, 'Active'),
        ('SKU-10121', 'Magnetic Sweeper Push 24-inch', 'Maintenance', 'Each', 115.00, 'Active'),
        ('SKU-10122', 'Sanding Disc 5-inch 80-Grit 50pk', 'Abrasives', 'Pack', 14.25, 'Active'),
        ('SKU-10123', 'Conveyor Roller 1.9in x 18in', 'Material Handling', 'Each', 12.50, 'Active'),
        ('SKU-10124', 'Pressure Washer 2000PSI Electric', 'Cleaning Supplies', 'Each', 299.00, 'Active'),
        ('SKU-10087', 'Lockout Tagout Station 6-Lock', 'Safety Equipment', 'Each', 185.00, 'Active'),  # DUPLICATE row 130
        ('SKU-10125', 'Floor Sign Wet Floor Yellow', 'Safety Equipment', 'Each', 12.50, 'Active'),
        ('SKU-10126', 'Grinding Disc 7-inch Depressed', 'Abrasives', 'Each', 4.75, 'Active'),
        ('SKU-10127', 'Drum Trolley Spill Containment', 'Environmental', 'Each', 345.00, 'Active'),
        ('SKU-10128', 'Steel Shelving Starter Unit 36x18', 'Storage', 'Each', 185.00, 'Active'),
        ('SKU-10129', 'Capacitor Start Motor 1/2HP', 'Electrical', 'Each', 89.00, 'Active'),
        ('SKU-10130', 'Vinyl Floor Tape 2-inch Yellow', 'Safety Equipment', 'Roll', 16.00, 'Active'),
        ('SKU-10131', 'Bolt Cutter 18-inch', 'Hand Tools', 'Each', 32.50, 'Active'),
        ('SKU-10132', 'Transfer Pump 12V DC Auto', 'Fluid Handling', 'Each', 55.00, 'Active'),
        ('SKU-10133', 'Hardhat Full Brim Yellow', 'Safety Equipment', 'Each', 18.00, 'Active'),
        ('SKU-10104', 'First Aid Kit 100-Piece Wall', 'Safety Equipment', 'Each', 65.00, 'Active'),  # DUPLICATE row 140
        ('SKU-10134', 'Pipe Thread Sealant PTFE Tape', 'Plumbing', 'Roll', 2.25, 'Active'),
        ('SKU-10135', 'Cable Gland Stainless M20 10pk', 'Electrical', 'Pack', 21.50, 'Active'),
        ('SKU-10136', 'Drum Pump Manual Rotary', 'Fluid Handling', 'Each', 48.00, 'Active'),
        ('SKU-10137', 'Scaffold Caster Wheel 8-inch Locking', 'Construction', 'Each', 35.00, 'Active'),
        ('SKU-10029', 'High-Visibility Safety Vest', 'Safety Equipment', 'Each', 8.75, 'Active'),  # DUPLICATE row 148
    ]

    # Write data rows starting at row 2
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Column G is left empty (no header, no data)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Total data rows: {len(data)}')
    print('Duplicate SKU rows (1-indexed data rows):')
    seen = {}
    for i, row in enumerate(data, 1):
        sku = row[0]
        if sku in seen:
            print(f'  Row {i+1} (data row {i}): {sku} (first at data row {seen[sku]})')
        else:
            seen[sku] = i


create_initial()
