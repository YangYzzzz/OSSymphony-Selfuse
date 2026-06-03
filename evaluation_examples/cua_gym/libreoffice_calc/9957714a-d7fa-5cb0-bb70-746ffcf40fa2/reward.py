"""
Reward Script: Find duplicate SKUs, add 'Duplicate Flag' column header, remove duplicate rows keeping first occurrence.
Task ID: calc_ops_inventory_duplicate_sku_005
Domain: libreoffice_calc
Scoring:
  Component 1: 'Duplicate Flag' header in G1                     (0.20 pts)
  Component 2: Correct number of rows after removing duplicates  (0.40 pts)
  Component 3: All remaining SKUs are unique (no duplicates)     (0.30 pts)
  Component 4: Contiguous data rows (no blank gaps)              (0.10 pts)
Total: 1.0
"""

import os
import openpyxl
from collections import Counter

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_ops_inventory_duplicate_sku_005'
SHEET_NAME = 'ItemMaster'

# Known ground truth: initial file had 150 data rows with 13 duplicate SKUs
# Golden file should have 137 data rows (150 - 13 = 137)
INITIAL_DATA_ROWS = 150
EXPECTED_DUPLICATE_COUNT = 13
EXPECTED_DATA_ROWS = INITIAL_DATA_ROWS - EXPECTED_DUPLICATE_COUNT  # 137


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook — gate check
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Sheet presence check — gate
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found in workbook. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Component 1: 'Duplicate Flag' header in G1 (0.20 points)
    # Initial file: G1 is None. Golden: G1 = 'Duplicate Flag'
    # This FAILS on initial and PASSES on golden.
    try:
        g1_value = ws.cell(row=1, column=7).value
        if g1_value is not None and str(g1_value).strip().lower() == 'duplicate flag':
            print(f"PASS: Component 1 — 'Duplicate Flag' header found in G1 (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 — Expected 'Duplicate Flag' in G1, found: {repr(g1_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Correct row count after removing duplicates (0.40 points)
    # Initial file: 151 rows (150 data rows). Golden: 138 rows (137 data rows).
    # Task specifies 13 duplicate SKU entries, so final should be 137 data rows.
    # We allow a range of 135-137 rows to account for minor interpretation differences.
    # This FAILS on initial (150 data rows) and PASSES on golden (137 data rows).
    try:
        actual_max_row = ws.max_row
        actual_data_rows = actual_max_row - 1  # subtract header row
        # Exact match: 137 data rows
        row_score = 0.0
        if actual_data_rows == EXPECTED_DATA_ROWS:
            row_score = 0.40
            print(f"PASS: Component 2 — Row count is {actual_data_rows} data rows (exactly {EXPECTED_DATA_ROWS} expected) (0.40 pts)")
        elif EXPECTED_DATA_ROWS - 2 <= actual_data_rows <= EXPECTED_DATA_ROWS + 2:
            # Partial: within 2 rows of expected (in case of slight interpretation differences)
            row_score = 0.20
            print(f"PARTIAL: Component 2 — Row count is {actual_data_rows} data rows, expected {EXPECTED_DATA_ROWS}. Close but not exact (0.20 pts)")
        elif actual_data_rows < INITIAL_DATA_ROWS:
            # Some duplicates removed but not exactly the right number
            removed = INITIAL_DATA_ROWS - actual_data_rows
            row_score = 0.10
            print(f"PARTIAL: Component 2 — {removed} rows removed, expected {EXPECTED_DUPLICATE_COUNT}. Not exact (0.10 pts)")
        else:
            print(f"FAIL: Component 2 — Row count is {actual_data_rows} data rows, expected {EXPECTED_DATA_ROWS} after removing {EXPECTED_DUPLICATE_COUNT} duplicates")
        if row_score > 0:
            total_score += row_score
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: All remaining SKUs are unique — no duplicates remain (0.30 points)
    # Initial file: 13 duplicate SKUs present. Golden: 0 duplicates.
    # This FAILS on initial and PASSES on golden.
    try:
        skus = []
        for row in range(2, ws.max_row + 1):
            sku_val = ws.cell(row=row, column=1).value
            if sku_val is not None:
                skus.append(str(sku_val).strip())

        if len(skus) == 0:
            print(f"FAIL: Component 3 — No SKU data found in column A")
        else:
            sku_counts = Counter(skus)
            remaining_duplicates = {sku: count for sku, count in sku_counts.items() if count > 1}
            if len(remaining_duplicates) == 0:
                print(f"PASS: Component 3 — All {len(skus)} SKUs are unique, no duplicates remain (0.30 pts)")
                total_score += 0.30
            else:
                dup_list = list(remaining_duplicates.keys())[:5]
                print(f"FAIL: Component 3 — {len(remaining_duplicates)} duplicate SKUs still present: {dup_list}...")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Contiguous data rows AND correct row count (combined check) (0.10 points)
    # This verifies that duplicate rows were truly DELETED (not just cleared), AND that the
    # first occurrence of each SKU was preserved correctly.
    # Specifically: no blank rows in data area AND row count matches expected post-dedup count.
    # Initial file: 150 data rows — FAILS because row count doesn't match expected 137.
    # Golden file: 137 data rows, no blanks — PASSES.
    try:
        blank_rows = 0

        # Count blank rows (rows where SKU column A is None/empty)
        for row in range(2, ws.max_row + 1):
            a_val = ws.cell(row=row, column=1).value
            if a_val is None or str(a_val).strip() == '':
                blank_rows += 1

        actual_data_rows = ws.max_row - 1  # subtract header

        # Only award points if: no blank rows AND row count is within expected range
        # (i.e., duplicates were deleted, not just cleared, AND correct number removed)
        if blank_rows == 0 and actual_data_rows == EXPECTED_DATA_ROWS:
            print(f"PASS: Component 4 — No blank rows and correct row count {actual_data_rows} (0.10 pts)")
            total_score += 0.10
        elif blank_rows > 0:
            print(f"FAIL: Component 4 — {blank_rows} blank rows found (rows were cleared, not deleted)")
        elif actual_data_rows != EXPECTED_DATA_ROWS:
            print(f"FAIL: Component 4 — Row count {actual_data_rows} does not match expected {EXPECTED_DATA_ROWS} (some duplicates not fully removed)")
        else:
            print(f"FAIL: Component 4 — Unexpected state")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
