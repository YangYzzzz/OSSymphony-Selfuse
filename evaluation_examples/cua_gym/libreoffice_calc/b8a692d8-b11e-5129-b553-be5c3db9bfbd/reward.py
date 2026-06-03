"""
Reward Script: Sort project status report by custom priority order
Task ID: calc_dop_sort_custom_007
Domain: libreoffice_calc
Scoring:
  - Component 1: Critical rows appear in rows 2-4 (0.4 pts)
  - Component 2: At Risk rows appear in rows 5-9 (0.3 pts)
  - Component 3: On Track rows appear in rows 10-17 (0.2 pts)
  - Component 4: Completed rows appear in rows 18-21 (0.1 pts)
Total: 1.0

The task requires sorting the 'Projects' sheet by Status column using
a custom priority order: Critical > At Risk > On Track > Completed.
The initial file has alphabetical order (At Risk, Completed, Critical, On Track).
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_dop_sort_custom_007'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Projects' sheet must exist
    if 'Projects' not in wb.sheetnames:
        print("CRITICAL: 'Projects' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Projects']

    # Verify that headers are in row 1 as expected (precondition gate, not scored)
    header_d = ws.cell(row=1, column=4).value
    if header_d != 'Status':
        print(f"CRITICAL: Header in D1 is '{header_d}', expected 'Status'. File structure is wrong.")
        print("REWARD: 0.0")
        return 0.0

    # Read all Status values from rows 2-21
    statuses = []
    for row in range(2, 22):
        val = ws.cell(row=row, column=4).value
        statuses.append(val)

    print(f"Status values in rows 2-21: {statuses}")

    # Component 1: Critical rows appear first (rows 2-4) — 0.4 points
    # The first 3 data rows must all be 'Critical'
    try:
        critical_rows = statuses[0:3]  # rows 2, 3, 4
        if all(v == 'Critical' for v in critical_rows):
            print(f"PASS: Component 1 — All 3 'Critical' rows appear first (rows 2-4) (0.4 pts)")
            print(f"      Found: {critical_rows}")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — Expected ['Critical','Critical','Critical'] in rows 2-4, found: {critical_rows}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: At Risk rows appear next (rows 5-9) — 0.3 points
    # The next 5 data rows must all be 'At Risk'
    try:
        at_risk_rows = statuses[3:8]  # rows 5, 6, 7, 8, 9
        if all(v == 'At Risk' for v in at_risk_rows):
            print(f"PASS: Component 2 — All 5 'At Risk' rows appear in rows 5-9 (0.3 pts)")
            print(f"      Found: {at_risk_rows}")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — Expected 5x 'At Risk' in rows 5-9, found: {at_risk_rows}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: On Track rows appear next (rows 10-17) — 0.2 points
    # The next 8 data rows must all be 'On Track'
    try:
        on_track_rows = statuses[8:16]  # rows 10, 11, 12, 13, 14, 15, 16, 17
        if all(v == 'On Track' for v in on_track_rows):
            print(f"PASS: Component 3 — All 8 'On Track' rows appear in rows 10-17 (0.2 pts)")
            print(f"      Found: {on_track_rows}")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — Expected 8x 'On Track' in rows 10-17, found: {on_track_rows}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Completed rows appear last (rows 18-21) — 0.1 points
    # The last 4 data rows must all be 'Completed'
    try:
        completed_rows = statuses[16:20]  # rows 18, 19, 20, 21
        if all(v == 'Completed' for v in completed_rows):
            print(f"PASS: Component 4 — All 4 'Completed' rows appear last (rows 18-21) (0.1 pts)")
            print(f"      Found: {completed_rows}")
            total_score += 0.1
        else:
            print(f"FAIL: Component 4 — Expected 4x 'Completed' in rows 18-21, found: {completed_rows}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
