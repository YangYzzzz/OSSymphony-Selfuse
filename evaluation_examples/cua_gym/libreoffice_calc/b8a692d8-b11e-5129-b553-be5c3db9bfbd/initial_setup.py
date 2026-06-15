"""
Initial Setup: Sort project status report by custom priority order
Task ID: calc_dop_sort_custom_007
Domain: libreoffice_calc

Creates a 'Projects' sheet with 20 rows of data.
Data is sorted alphabetically by Status: At Risk, Completed, Critical, On Track
(NOT the desired custom priority order - that is what the agent must apply)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

WORKDIR = '/home/user'
TASK_ID = 'calc_dop_sort_custom_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Projects'

    # --- Headers in row 1 ---
    headers = ['Project ID', 'Project Name', 'Manager', 'Status', 'Due Date', 'Budget']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # --- Data rows sorted ALPHABETICALLY by Status (At Risk, Completed, Critical, On Track) ---
    # 5 At Risk rows (rows 2-6), 4 Completed rows (rows 7-10), 3 Critical rows (rows 11-13), 8 On Track rows (rows 14-21)
    data = [
        # At Risk (5 rows)
        ['PRJ-003', 'ERP System Upgrade',        'Rebecca Torres',    'At Risk',   date(2025, 6, 30),  320000],
        ['PRJ-007', 'Supply Chain Optimization',  'Daniel Kim',        'At Risk',   date(2025, 7, 15),  215000],
        ['PRJ-012', 'Customer Data Migration',    'Sandra Patel',      'At Risk',   date(2025, 5, 20),  175000],
        ['PRJ-016', 'Legacy System Decommission', 'Marcus Chen',       'At Risk',   date(2025, 8, 10),  290000],
        ['PRJ-019', 'Compliance Audit Prep',      'Linda Hoffman',     'At Risk',   date(2025, 6, 15),  130000],
        # Completed (4 rows)
        ['PRJ-001', 'Annual Report Automation',   'James Wilson',      'Completed', date(2025, 3, 31),   95000],
        ['PRJ-005', 'HR Portal Redesign',         'Angela Martinez',   'Completed', date(2025, 2, 28),  145000],
        ['PRJ-009', 'Vendor Management System',   'Thomas Brown',      'Completed', date(2025, 4, 15),  210000],
        ['PRJ-015', 'Data Warehouse Migration',   'Christine Lee',     'Completed', date(2025, 1, 31),  385000],
        # Critical (3 rows)
        ['PRJ-002', 'Cloud Infrastructure Move',  'Michael Scott',     'Critical',  date(2025, 5, 1),   450000],
        ['PRJ-008', 'Security Patch Rollout',     'Natasha Ivanova',   'Critical',  date(2025, 4, 25),   80000],
        ['PRJ-014', 'Payment Gateway Overhaul',   'Robert Garcia',     'Critical',  date(2025, 5, 10),  560000],
        # On Track (8 rows)
        ['PRJ-004', 'Mobile App Refresh',         'Emily Nguyen',      'On Track',  date(2025, 9, 30),  190000],
        ['PRJ-006', 'Business Intelligence Dash', 'Kevin O\'Brien',    'On Track',  date(2025, 10, 15), 265000],
        ['PRJ-010', 'Network Infrastructure',     'Sofia Rodriguez',   'On Track',  date(2025, 8, 31),  340000],
        ['PRJ-011', 'Employee Training Platform', 'Brian Thompson',    'On Track',  date(2025, 11, 30), 155000],
        ['PRJ-013', 'CRM Implementation',         'Vanessa Williams',  'On Track',  date(2025, 12, 15), 420000],
        ['PRJ-017', 'Digital Marketing Campaign', 'Jason Park',        'On Track',  date(2025, 7, 31),   72000],
        ['PRJ-018', 'Inventory Management V2',    'Priya Sharma',      'On Track',  date(2025, 9, 15),  198000],
        ['PRJ-020', 'AI-Assisted Reporting',      'David Okonkwo',     'On Track',  date(2025, 10, 31), 310000],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14

    # Format Due Date column as date
    for r in range(2, 22):
        ws.cell(row=r, column=5).number_format = 'yyyy-mm-dd'

    # Format Budget column as currency
    for r in range(2, 22):
        ws.cell(row=r, column=6).number_format = '$#,##0'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet: Projects, 20 data rows (rows 2-21)')
    print(f'Status distribution: 5 At Risk, 4 Completed, 3 Critical, 8 On Track')
    print(f'Sorted alphabetically by Status (NOT the custom priority order)')


create_initial()
