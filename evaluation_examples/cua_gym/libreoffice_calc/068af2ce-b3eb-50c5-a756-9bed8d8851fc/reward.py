"""
Reward Script: Multi-sheet workbook with department tabs and Summary sheet formulas
Task ID: calc_hr_046
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Engineering SUM formula in Summary!B2
  Component 2 (0.25): Sales SUM formula in Summary!B3
  Component 3 (0.25): HR SUM formula in Summary!B4
  Component 4 (0.25): Grand Total SUM formula in Summary!B5
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_046'


def check_formula(ws, coord, expected_formula):
    """Check if cell contains expected formula (case-insensitive, space-stripped)."""
    val = ws[coord].value
    if not isinstance(val, str):
        return False, val
    return (val.upper().replace(" ", "") == expected_formula.upper().replace(" ", "")), val


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Summary sheet must exist
    if 'Summary' not in wb.sheetnames:
        print("FAIL: 'Summary' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Summary']

    # Component 1: Engineering SUM formula in Summary!B2 (0.25 points)
    try:
        match, actual = check_formula(ws, 'B2', '=SUM(Engineering.B2:B5)')
        if match:
            print(f"PASS: Component 1 — Engineering formula in B2: {actual} (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — expected =SUM(Engineering.B2:B5), found: {repr(actual)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Sales SUM formula in Summary!B3 (0.25 points)
    try:
        match, actual = check_formula(ws, 'B3', '=SUM(Sales.B2:B4)')
        if match:
            print(f"PASS: Component 2 — Sales formula in B3: {actual} (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — expected =SUM(Sales.B2:B4), found: {repr(actual)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: HR SUM formula in Summary!B4 (0.25 points)
    try:
        match, actual = check_formula(ws, 'B4', '=SUM(HR.B2:B4)')
        if match:
            print(f"PASS: Component 3 — HR formula in B4: {actual} (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 3 — expected =SUM(HR.B2:B4), found: {repr(actual)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Grand Total SUM formula in Summary!B5 (0.25 points)
    try:
        match, actual = check_formula(ws, 'B5', '=SUM(B2:B4)')
        if match:
            print(f"PASS: Component 4 — Grand Total formula in B5: {actual} (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 4 — expected =SUM(B2:B4), found: {repr(actual)}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
