"""
Initial Setup: Create multi-sheet HR workbook with department data
Task ID: calc_hr_046
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_046'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Engineering ---
    ws_eng = wb.active
    ws_eng.title = 'Engineering'
    ws_eng['A1'] = 'Employee'
    ws_eng['B1'] = 'Salary'
    eng_data = [
        ['Sarah Chen', 95000],
        ['Marcus Johnson', 88000],
        ['Priya Patel', 72000],
        ['David Kim', 65000],
    ]
    for r, row_data in enumerate(eng_data, 2):
        ws_eng.cell(row=r, column=1, value=row_data[0])
        ws_eng.cell(row=r, column=2, value=row_data[1])

    # --- Sheet 2: Sales ---
    ws_sales = wb.create_sheet('Sales')
    ws_sales['A1'] = 'Employee'
    ws_sales['B1'] = 'Salary'
    sales_data = [
        ['Jessica Rivera', 90000],
        ['Thomas Wright', 82000],
        ['Amanda Foster', 73000],
    ]
    for r, row_data in enumerate(sales_data, 2):
        ws_sales.cell(row=r, column=1, value=row_data[0])
        ws_sales.cell(row=r, column=2, value=row_data[1])

    # --- Sheet 3: HR ---
    ws_hr = wb.create_sheet('HR')
    ws_hr['A1'] = 'Employee'
    ws_hr['B1'] = 'Salary'
    hr_data = [
        ['Laura Mitchell', 68000],
        ['Robert Chang', 58000],
        ['Emily Santos', 54000],
    ]
    for r, row_data in enumerate(hr_data, 2):
        ws_hr.cell(row=r, column=1, value=row_data[0])
        ws_hr.cell(row=r, column=2, value=row_data[1])

    # --- Sheet 4: Summary (labels only, NO formulas) ---
    ws_summary = wb.create_sheet('Summary')
    ws_summary['A1'] = 'Department'
    ws_summary['B1'] = 'Total Salary'
    ws_summary['A2'] = 'Engineering'
    ws_summary['A3'] = 'Sales'
    ws_summary['A4'] = 'HR'
    ws_summary['A5'] = 'Grand Total'
    # B2:B5 left empty — the task is to add formulas here

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
