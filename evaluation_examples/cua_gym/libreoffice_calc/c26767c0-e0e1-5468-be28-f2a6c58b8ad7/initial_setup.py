"""
Initial Setup: Task Manager spreadsheet with due dates for conditional formatting task
Task ID: calc_fmt_condfmt_date_occurring_068
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_condfmt_date_occurring_068'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Task Manager ---
    ws = wb.active
    ws.title = 'Task Manager'

    # Headers
    headers = ['Task', 'Assignee', 'Priority', 'Due Date']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Task data with realistic content
    # Column D stores date serials (integer) with format 'General' — no date formatting
    # Today is approximately serial 46085 (2026-03-04)
    # Mix of past, current-week, next-7-days, and future dates
    data = [
        # (Task, Assignee, Priority, DateSerial)
        ('Update security certificates',      'Sarah Chen',       'High',   46082),  # 2026-03-01 (past)
        ('Quarterly financial report',        'Marcus Johnson',   'High',   46086),  # 2026-03-05 (next 7 days)
        ('Deploy v2.4 hotfix',                'Elena Rodriguez',  'Critical', 46087),  # 2026-03-06 (next 7 days)
        ('Client onboarding — Acme Corp',     'David Kim',        'Medium', 46055),  # 2026-02-02 (past)
        ('Database backup migration',         'Priya Sharma',     'High',   46088),  # 2026-03-07 (next 7 days)
        ('Design review for dashboard',       'Tom Nguyen',       'Low',    46115),  # 2026-04-03 (future)
        ('Fix login timeout bug',             'Fatima Al-Hassan', 'Critical', 46089), # 2026-03-08 (next 7 days)
        ('Write API documentation',           'James Liu',        'Low',    46099),  # 2026-03-18 (future)
        ('Conduct team performance reviews',  'Sarah Chen',       'Medium', 46071),  # 2026-02-18 (past)
        ('Prepare demo for investor meeting', 'Marcus Johnson',   'High',   46090),  # 2026-03-09 (next 7 days)
        ('Refactor authentication module',    'Elena Rodriguez',  'Medium', 46106),  # 2026-03-25 (future)
        ('Renew software licenses',           'David Kim',        'High',   46091),  # 2026-03-10 (next 7 days)
        ('Update employee handbook',          'Priya Sharma',     'Low',    46130),  # future
        ('Fix data export CSV bug',           'Tom Nguyen',       'Medium', 46085),  # 2026-03-04 (today)
        ('Launch email campaign Q1',          'Fatima Al-Hassan', 'High',   46092),  # 2026-03-11 (next 7 days)
        ('Performance testing — v3.0',        'James Liu',        'Medium', 46120),  # future
        ('Migrate servers to new datacenter', 'Sarah Chen',       'Critical', 46078), # 2026-02-25 (past)
        ('Implement GDPR compliance checks',  'Marcus Johnson',   'High',   46093),  # 2026-03-12 (future)
        ('Schedule training for new hires',   'Elena Rodriguez',  'Low',    46150),  # future
        ('Audit access logs for Q4',          'David Kim',        'Medium', 46060),  # past
        ('Build automated test suite',        'Priya Sharma',     'Medium', 46094),  # 2026-03-13 (future)
        ('Update privacy policy',             'Tom Nguyen',       'High',   46086),  # 2026-03-05 (next 7 days)
        ('Resolve DNS propagation issues',    'Fatima Al-Hassan', 'Critical', 46083), # 2026-03-02 (past)
        ('Integrate Stripe payment gateway',  'James Liu',        'High',   46095),  # 2026-03-14 (future)
        ('Create Q1 sales dashboard',         'Sarah Chen',       'Medium', 46087),  # 2026-03-06 (next 7 days)
        ('Organize IT assets inventory',      'Marcus Johnson',   'Low',    46140),  # future
        ('Coordinate office relocation',      'Elena Rodriguez',  'High',   46088),  # 2026-03-07 (next 7 days)
        ('Review vendor contracts',           'David Kim',        'Medium', 46100),  # future
        ('Set up CI/CD pipeline',             'Priya Sharma',     'High',   46089),  # 2026-03-08 (next 7 days)
        ('Update firewall rules',             'Tom Nguyen',       'Critical', 46084), # 2026-03-03 (past)
        ('Onboard new DevOps engineer',       'Fatima Al-Hassan', 'Medium', 46110),  # future
        ('Archive old project files',         'James Liu',        'Low',    46065),  # past
        ('Implement dark mode UI',            'Sarah Chen',       'Low',    46125),  # future
        ('Fix memory leak in analytics svc',  'Marcus Johnson',   'Critical', 46090), # 2026-03-09 (next 7 days)
        ('Prepare board meeting slides',      'Elena Rodriguez',  'High',   46091),  # 2026-03-10 (next 7 days)
        ('Update SSL certificates — prod',    'David Kim',        'High',   46086),  # 2026-03-05 (next 7 days)
        ('Deploy mobile app v1.5',            'Priya Sharma',     'Medium', 46135),  # future
        ('Complete SOC2 audit prep',          'Tom Nguyen',       'High',   46092),  # 2026-03-11 (next 7 days)
        ('Run load tests for v3.0 release',   'Fatima Al-Hassan', 'Medium', 46096),  # future
    ]

    for r, row_data in enumerate(data, 2):
        task, assignee, priority, date_serial = row_data
        ws.cell(row=r, column=1, value=task)
        ws.cell(row=r, column=2, value=assignee)
        ws.cell(row=r, column=3, value=priority)
        # Store date as integer serial with 'General' format (no date formatting)
        cell = ws.cell(row=r, column=4, value=date_serial)
        cell.number_format = 'General'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12

    # NO conditional formatting — task requires adding it from scratch

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet: Task Manager, Rows: 1 header + 39 data rows')
    print(f'Column D: date serials (General format, no conditional formatting)')


create_initial()
