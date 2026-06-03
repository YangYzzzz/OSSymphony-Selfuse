"""
Reward Script: Add conditional formatting for 'next 7 days' due dates
Task ID: calc_fmt_condfmt_date_occurring_068
Domain: libreoffice_calc

Scoring Rubric:
  Component 1 (0.30): CF rule applied to D2:D40 range
  Component 2 (0.30): Rule type is timePeriod (date occurring rule) with a 'next week' or equivalent period
  Component 3 (0.25): Background fill color is yellow #FFEB9C (ARGB: FFFFEB9C)
  Component 4 (0.15): Font color is #9C6500 (ARGB: FF9C6500)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmt_condfmt_date_occurring_068'
SHEET_NAME = 'Task Manager'
CF_RANGE = 'D2:D40'


def normalize_color(color_str):
    """Normalize color to 8-char uppercase ARGB. Handle both 6-char and 8-char."""
    if not color_str:
        return None
    s = str(color_str).upper().strip()
    if len(s) == 6:
        s = 'FF' + s
    return s


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook (precondition gate)
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Sheet existence check (precondition gate — NOT scored)
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Gather all conditional formatting rules
    all_cfs = list(ws.conditional_formatting)

    # ---------------------------------------------------------
    # Component 1: CF rule applied to D2:D40 range (0.30 points)
    # This FAILS on initial (no CF at all) and PASSES on golden.
    # ---------------------------------------------------------
    target_rule = None
    target_cf_obj = None
    try:
        # Look for a ConditionalFormatting entry covering D2:D40
        for cf in all_cfs:
            cf_str = str(cf)
            # The range should reference D2:D40 (the exact range or superset)
            if 'D2:D40' in cf_str or 'D2' in cf_str:
                if len(cf.rules) > 0:
                    target_cf_obj = cf
                    target_rule = cf.rules[0]
                    break

        if target_rule is not None:
            print(f"PASS: Component 1 — CF rule found covering D2:D40 range (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 — No CF rule found on D2:D40. "
                  f"Existing CF ranges: {[str(c) for c in all_cfs]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ---------------------------------------------------------
    # Component 2: Rule type is timePeriod (date occurring rule) with
    # a 'next 7 days' / 'next week' type period (0.30 points)
    # This FAILS on initial (no CF) and PASSES on golden.
    # ---------------------------------------------------------
    try:
        if target_rule is not None:
            rule_type = getattr(target_rule, 'type', None)
            time_period = getattr(target_rule, 'timePeriod', None)

            # Acceptable time periods for 'next 7 days' / 'upcoming' dates:
            # 'nextWeek', 'next7Days', 'tomorrow'
            VALID_TIME_PERIODS = {'nextweek', 'next7days', 'tomorrow', 'thisweek'}

            if rule_type == 'timePeriod':
                period_lower = str(time_period).lower() if time_period else ''
                if period_lower in VALID_TIME_PERIODS:
                    print(f"PASS: Component 2 — Rule is timePeriod with period='{time_period}' (0.30 pts)")
                    total_score += 0.30
                else:
                    print(f"FAIL: Component 2 — Rule is timePeriod but period='{time_period}'. "
                          f"Expected one of: nextWeek, next7Days, tomorrow")
            else:
                # Also accept formula-based rules that compute next 7 days
                formula = getattr(target_rule, 'formula', None)
                if rule_type == 'expression' or rule_type == 'formula':
                    # A formula that involves TODAY() and a 7-day window is acceptable
                    formula_str = str(formula).upper() if formula else ''
                    if 'TODAY' in formula_str and ('7' in formula_str or 'WEEK' in formula_str):
                        print(f"PASS: Component 2 — Formula-based date rule with TODAY() and 7-day window (0.30 pts)")
                        total_score += 0.30
                    else:
                        print(f"FAIL: Component 2 — Rule type='{rule_type}', formula='{formula}'. "
                              f"Expected timePeriod or formula with TODAY()+7 logic")
                else:
                    print(f"FAIL: Component 2 — Rule type='{rule_type}'. Expected 'timePeriod'")
        else:
            print("FAIL: Component 2 — No target rule found (prerequisite for Component 1 failed)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---------------------------------------------------------
    # Component 3: Background fill color is yellow #FFEB9C (0.25 points)
    # ARGB: FFFFEB9C
    # This FAILS on initial (no CF) and PASSES on golden.
    # ---------------------------------------------------------
    EXPECTED_BG = 'FFFFEB9C'
    try:
        if target_rule is not None:
            dxf = getattr(target_rule, 'dxf', None)
            if dxf is not None and dxf.fill is not None:
                try:
                    bg_rgb = dxf.fill.fgColor.rgb
                    bg_norm = normalize_color(bg_rgb)
                    # Also accept the 6-char variant FFEB9C
                    if bg_norm in (EXPECTED_BG, 'FF' + 'FFEB9C'):
                        print(f"PASS: Component 3 — Background fill color is #{bg_rgb} (0.25 pts)")
                        total_score += 0.25
                    else:
                        print(f"FAIL: Component 3 — Background fill color is #{bg_rgb}, "
                              f"expected #{EXPECTED_BG} (#FFEB9C yellow)")
                except Exception as e2:
                    print(f"FAIL: Component 3 — Could not read fgColor.rgb: {e2}")
            else:
                print("FAIL: Component 3 — No dxf fill found on rule")
        else:
            print("FAIL: Component 3 — No target rule found (prerequisite for Component 1 failed)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ---------------------------------------------------------
    # Component 4: Font color is #9C6500 (0.15 points)
    # ARGB: FF9C6500
    # This FAILS on initial (no CF) and PASSES on golden.
    # ---------------------------------------------------------
    EXPECTED_FONT = 'FF9C6500'
    try:
        if target_rule is not None:
            dxf = getattr(target_rule, 'dxf', None)
            if dxf is not None and dxf.font is not None:
                try:
                    font_rgb = dxf.font.color.rgb
                    font_norm = normalize_color(font_rgb)
                    if font_norm == EXPECTED_FONT or font_norm == '9C6500':
                        print(f"PASS: Component 4 — Font color is #{font_rgb} (0.15 pts)")
                        total_score += 0.15
                    else:
                        print(f"FAIL: Component 4 — Font color is #{font_rgb}, "
                              f"expected #{EXPECTED_FONT} (#9C6500 dark orange)")
                except Exception as e2:
                    print(f"FAIL: Component 4 — Could not read font color.rgb: {e2}")
            else:
                print("FAIL: Component 4 — No dxf font found on rule")
        else:
            print("FAIL: Component 4 — No target rule found (prerequisite for Component 1 failed)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
