"""
Reward Script: Add running total column with conditional alerts for budget threshold breaches
Task ID: calc_gpm_030
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Running total formulas in D2:D13
  Component 2 (0.20): Budget remaining formulas in E2:E13
  Component 3 (0.15): Alert formulas in F2:F13
  Component 4 (0.15): Conditional formatting on F column (3-level alerts)
  Component 5 (0.10): Conditional formatting on E column (color-coded text)
  Component 6 (0.05): Color scale on D2:D13
  Component 7 (0.10): Borders on all data cells
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_030'


def get_cf_sqref(cf_obj):
    """Extract the cell range string from a ConditionalFormatting object."""
    try:
        return str(cf_obj.sqref)
    except Exception:
        pass
    try:
        return ' '.join(str(r) for r in cf_obj.cells.ranges)
    except Exception:
        return ''


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists
    if 'Expenses' not in wb.sheetnames:
        print("FAIL: 'Expenses' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Expenses']

    # Component 1: Running total formulas in D2:D13 (0.25 points)
    try:
        d_pass = 0
        # D2 should be =C2 (first running total)
        d2_val = ws.cell(row=2, column=4).value
        if d2_val and isinstance(d2_val, str):
            d2_norm = d2_val.upper().replace(" ", "")
            if d2_norm == "=C2":
                d_pass += 1
        # D3:D13 should be =D(prev)+C(current)
        for r in range(3, 14):
            cell_val = ws.cell(row=r, column=4).value
            if cell_val and isinstance(cell_val, str):
                expected = f"=D{r-1}+C{r}".upper()
                actual = cell_val.upper().replace(" ", "")
                if actual == expected:
                    d_pass += 1
        if d_pass == 12:
            print(f"PASS: Component 1 -- All 12 running total formulas correct in D2:D13 (0.25 pts)")
            total_score += 0.25
        elif d_pass >= 6:
            partial = round(0.25 * d_pass / 12, 2)
            print(f"PARTIAL: Component 1 -- {d_pass}/12 running total formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 -- Only {d_pass}/12 running total formulas found in D2:D13")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Budget remaining formulas in E2:E13 (0.20 points)
    try:
        e_pass = 0
        for r in range(2, 14):
            cell_val = ws.cell(row=r, column=5).value
            if cell_val and isinstance(cell_val, str):
                actual = cell_val.upper().replace(" ", "")
                # Accept =$H$1-D{r} or =H1-D{r}
                valid_forms = [
                    f"=$H$1-D{r}".upper(),
                    f"=H1-D{r}".upper(),
                ]
                if actual in valid_forms:
                    e_pass += 1
        if e_pass == 12:
            print(f"PASS: Component 2 -- All 12 budget remaining formulas correct in E2:E13 (0.20 pts)")
            total_score += 0.20
        elif e_pass >= 6:
            partial = round(0.20 * e_pass / 12, 2)
            print(f"PARTIAL: Component 2 -- {e_pass}/12 budget remaining formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- Only {e_pass}/12 budget remaining formulas found in E2:E13")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Alert formulas in F2:F13 (0.15 points)
    try:
        f_pass = 0
        for r in range(2, 14):
            cell_val = ws.cell(row=r, column=6).value
            if cell_val and isinstance(cell_val, str):
                actual = cell_val.upper().replace(" ", "")
                # The IF formula checks E<=0 for OVER BUDGET, E<2000 for WARNING, else OK
                expected = f'=IF(E{r}<=0,"OVERBUDGET",IF(E{r}<2000,"WARNING","OK"))'.upper().replace(" ", "")
                if actual == expected:
                    f_pass += 1
        if f_pass == 12:
            print(f"PASS: Component 3 -- All 12 alert formulas correct in F2:F13 (0.15 pts)")
            total_score += 0.15
        elif f_pass >= 6:
            partial = round(0.15 * f_pass / 12, 2)
            print(f"PARTIAL: Component 3 -- {f_pass}/12 alert formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 -- Only {f_pass}/12 alert formulas found in F2:F13")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Conditional formatting on F column - 3-level alert system (0.15 points)
    try:
        has_over_budget_rule = False
        has_warning_rule = False
        has_ok_rule = False

        for cf in ws.conditional_formatting:
            sqref = get_cf_sqref(cf)
            # Only process CF ranges that start with F (like "F2:F13")
            if not sqref.startswith('F'):
                continue
            for rule in cf.rules:
                if rule.type == 'cellIs' and rule.operator == 'equal':
                    formula = getattr(rule, 'formula', [])
                    if formula:
                        f_str = str(formula[0]).upper().replace(" ", "")
                        if 'OVERBUDGET' in f_str or 'OVER BUDGET' in f_str:
                            if rule.dxf and rule.dxf.fill and rule.dxf.fill.fgColor:
                                try:
                                    fill_rgb = rule.dxf.fill.fgColor.rgb
                                    if fill_rgb and 'FF0000' in fill_rgb:
                                        has_over_budget_rule = True
                                except Exception:
                                    pass
                        elif 'WARNING' in f_str:
                            if rule.dxf and rule.dxf.fill and rule.dxf.fill.fgColor:
                                try:
                                    fill_rgb = rule.dxf.fill.fgColor.rgb
                                    if fill_rgb and 'FFFF00' in fill_rgb:
                                        has_warning_rule = True
                                except Exception:
                                    pass
                        elif '"OK"' in f_str:
                            if rule.dxf and rule.dxf.fill and rule.dxf.fill.fgColor:
                                try:
                                    fill_rgb = rule.dxf.fill.fgColor.rgb
                                    if fill_rgb and '00FF00' in fill_rgb:
                                        has_ok_rule = True
                                except Exception:
                                    pass

        cf_f_count = sum([has_over_budget_rule, has_warning_rule, has_ok_rule])
        if cf_f_count == 3:
            print(f"PASS: Component 4 -- All 3 conditional formatting rules on F column (0.15 pts)")
            total_score += 0.15
        elif cf_f_count >= 1:
            partial = round(0.15 * cf_f_count / 3, 2)
            print(f"PARTIAL: Component 4 -- {cf_f_count}/3 CF rules on F column ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 -- No CF rules on F column (OVER_BUDGET={has_over_budget_rule}, WARNING={has_warning_rule}, OK={has_ok_rule})")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # Component 5: Conditional formatting on E column - color-coded text (0.10 points)
    try:
        has_red_text = False
        has_orange_text = False
        has_green_text = False

        for cf in ws.conditional_formatting:
            sqref = get_cf_sqref(cf)
            # Only process CF ranges that start with E (like "E2:E13")
            if not sqref.startswith('E'):
                continue
            for rule in cf.rules:
                if rule.type == 'cellIs' and rule.dxf and rule.dxf.font and rule.dxf.font.color:
                    try:
                        font_rgb = rule.dxf.font.color.rgb
                        if not font_rgb:
                            continue
                        # Red text for < 0
                        if 'FF0000' in font_rgb and rule.operator == 'lessThan':
                            has_red_text = True
                        # Orange text for 0 to 2000
                        elif ('FF8C00' in font_rgb or 'FFA500' in font_rgb) and rule.operator == 'between':
                            has_orange_text = True
                        # Green text for >= 2000
                        elif ('008000' in font_rgb or '00FF00' in font_rgb) and rule.operator in ('greaterThanOrEqual', 'greaterThan'):
                            has_green_text = True
                    except Exception:
                        pass

        e_cf_count = sum([has_red_text, has_orange_text, has_green_text])
        if e_cf_count == 3:
            print(f"PASS: Component 5 -- All 3 conditional formatting rules on E column (0.10 pts)")
            total_score += 0.10
        elif e_cf_count >= 1:
            partial = round(0.10 * e_cf_count / 3, 2)
            print(f"PARTIAL: Component 5 -- {e_cf_count}/3 CF rules on E column ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 -- No CF on E column (red={has_red_text}, orange={has_orange_text}, green={has_green_text})")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    # Component 6: Color scale (gradient green to red) on D2:D13 (0.05 points)
    try:
        has_color_scale = False
        for cf in ws.conditional_formatting:
            sqref = get_cf_sqref(cf)
            if sqref.startswith('D'):
                for rule in cf.rules:
                    if rule.type == 'colorScale':
                        has_color_scale = True
                        break
            if has_color_scale:
                break

        if has_color_scale:
            print(f"PASS: Component 6 -- Color scale found on D column (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 -- No color scale conditional formatting found on D column")
    except Exception as e:
        print(f"ERROR: Component 6 -- {e}")

    # Component 7: Borders on all data cells (0.10 points)
    try:
        border_pass = 0
        border_total = 0
        # Check borders on key cells across the data range
        check_cells = [(1, 1), (1, 4), (1, 6), (2, 1), (2, 4), (2, 6),
                       (7, 1), (7, 4), (7, 6), (13, 1), (13, 4), (13, 6)]
        for r, c in check_cells:
            border_total += 1
            cell = ws.cell(row=r, column=c)
            if (cell.border.left.style and cell.border.right.style and
                    cell.border.top.style and cell.border.bottom.style):
                border_pass += 1

        if border_pass == border_total:
            print(f"PASS: Component 7 -- All {border_total} sampled cells have borders (0.10 pts)")
            total_score += 0.10
        elif border_pass >= border_total // 2:
            partial = round(0.10 * border_pass / border_total, 2)
            print(f"PARTIAL: Component 7 -- {border_pass}/{border_total} sampled cells have borders ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 7 -- Only {border_pass}/{border_total} sampled cells have borders")
    except Exception as e:
        print(f"ERROR: Component 7 -- {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
