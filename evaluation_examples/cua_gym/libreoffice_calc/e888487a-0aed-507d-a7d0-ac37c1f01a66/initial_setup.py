"""
Initial Setup: Expense tracker with raw data for running total task
Task ID: calc_gpm_030
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_030'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"

    # --- Header row styling ---
    header_fill = PatternFill(start_color="FFCC6600", end_color="FFCC6600", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")

    headers = ["Date", "Description", "Amount", "Running Total", "Budget Remaining", "Alert"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # G1: Total Budget label, H1: budget value
    g1 = ws.cell(row=1, column=7, value="Total Budget")
    g1.font = Font(bold=True)
    h1 = ws.cell(row=1, column=8, value=15000)
    h1.number_format = '$#,##0.00'
    h1.font = Font(bold=True)

    # --- Expense data rows 2-13 (12 entries) ---
    expenses = [
        ["2025-09-01", "Venue deposit", 3000],
        ["2025-09-05", "Catering advance", 5000],
        ["2025-09-10", "Flowers", 1200],
        ["2025-09-14", "DJ deposit", 500],
        ["2025-09-18", "Invitations printing", 350],
        ["2025-09-22", "Photography", 2000],
        ["2025-09-28", "Linen rental", 800],
        ["2025-10-02", "Centerpieces", 600],
        ["2025-10-06", "Valet parking", 400],
        ["2025-10-10", "AV equipment", 750],
        ["2025-10-15", "Favors", 300],
        ["2025-10-20", "Miscellaneous", 400],
    ]

    for r, (date_str, desc, amount) in enumerate(expenses, 2):
        ws.cell(row=r, column=1, value=date_str)
        ws.cell(row=r, column=2, value=desc)
        amt_cell = ws.cell(row=r, column=3, value=amount)
        amt_cell.number_format = '$#,##0.00'

    # Columns D, E, F are intentionally left EMPTY — task is to add formulas there

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
