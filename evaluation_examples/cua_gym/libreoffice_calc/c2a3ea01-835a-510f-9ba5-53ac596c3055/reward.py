"""
FINAL REWARD SCRIPT - SUCCESS
Task: I want the monthly total downloads calculated in a new row called "Total Downloads", then make a line chart with months on the x-axis showing the trend.
Generated: 2025-11-24 07:37:22
Status: success
Model: o3
Total Steps: 6
"""

import openpyxl
import math
import re
from openpyxl.chart.line_chart import LineChart
from openpyxl.utils import get_column_letter

# -----------------------------------------------------------------------------
# Reward Script for: “Calculate a new row called ‘Total Downloads’ containing
# monthly sums and create a line chart with months on the x-axis showing the
# trend.”
# -----------------------------------------------------------------------------
# Scoring rubric (max 1.0):
#  0.35 – Row labelled “Total Downloads” exists
#  0.25 – Row contains correct SUM formulas / numeric totals
#  0.20 – A line chart exists in the workbook
#  0.20 – Chart series/categories reference the header-row months and the
#          “Total Downloads” row values
# -----------------------------------------------------------------------------

FILE_PATH = (
    "/home/user/"
    "i_want_the_monthly_total_downloads_calculated_in_a_new_row_called_total_"
    "downloads_then_make_a_line_c.xlsx"
)

MONTH_ABBR = [
    "jan",
    "feb",
    "mar",
    "apr",
    "may",
    "jun",
    "jul",
    "aug",
    "sep",
    "oct",
    "nov",
    "dec",
]


# ------------------------------ helper functions -----------------------------

def parse_rows_from_ref(ref: str):
    """Return start & end row numbers mentioned in an Excel range reference."""
    if not ref:
        return []
    # remove sheet part before '!'
    if "!" in ref:
        ref = ref.split("!")[1]
    start, end = (ref.split(":") + [ref])[:2]  # handle single-cell refs too
    rows = []
    for addr in (start, end):
        m = re.search(r"(\d+)$", addr)
        rows.append(int(m.group(1)) if m else None)
    return rows


def locate_header_and_totals(sheet):
    header_row = None
    totals_row = None
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        first = row[0]
        if isinstance(first, str):
            if first.strip().lower() == "platform":
                header_row = idx
            if first.strip().lower() == "total downloads":
                totals_row = idx
    return header_row, totals_row


def month_columns(sheet, header_row):
    cols = []
    for cell in sheet[header_row]:
        if cell.column == 1:
            continue
        if (
            isinstance(cell.value, str)
            and cell.value.strip().lower()[:3] in MONTH_ABBR
        ):
            cols.append(cell.column)
    return cols


def formulas_or_values_ok(sheet, header_row, totals_row, month_cols):
    data_start = header_row + 1
    data_end = totals_row - 1
    if data_end < data_start:
        return False
    for col in month_cols:
        cell = sheet.cell(row=totals_row, column=col)
        val = cell.value
        # ---------- Formula path ----------
        if isinstance(val, str) and val.strip().startswith("="):
            m = re.match(r"=\s*SUM\(([^)]+)\)", val.strip(), re.IGNORECASE)
            if not m:
                return False
            rng = m.group(1).replace("$", "")
            if ":" not in rng:
                return False
            start_ref, end_ref = rng.split(":")
            # Column letters must match expected column
            expected_col_letter = get_column_letter(col).upper()
            if (
                re.sub(r"\d", "", start_ref).upper() != expected_col_letter
                or re.sub(r"\d", "", end_ref).upper() != expected_col_letter
            ):
                return False
            # Row numbers must match data section
            try:
                s_row = int(re.findall(r"\d+", start_ref)[0])
                e_row = int(re.findall(r"\d+", end_ref)[0])
            except (IndexError, ValueError):
                return False
            if s_row != data_start or e_row != data_end:
                return False
        # ---------- Static value path ----------
        else:
            try:
                numeric_val = float(val)
            except (TypeError, ValueError):
                return False
            expected = 0.0
            for r in range(data_start, data_end + 1):
                try:
                    expected += float(sheet.cell(row=r, column=col).value)
                except (TypeError, ValueError):
                    return False
            if not math.isclose(expected, numeric_val, rel_tol=1e-3):
                return False
    return True


def locate_line_chart(wb, header_row, totals_row):
    found = False
    refs_ok = False
    for ws in wb.worksheets:
        charts = getattr(ws, "_charts", []) or getattr(ws, "charts", [])
        for ch in charts:
            if isinstance(ch, LineChart):
                found = True
                for ser in ch.series:
                    cat_ref = getattr(getattr(ser, "cat", None), "numRef", None)
                    val_ref = getattr(getattr(ser, "val", None), "numRef", None)
                    cat_f = getattr(cat_ref, "f", None)
                    val_f = getattr(val_ref, "f", None)
                    cat_rows = parse_rows_from_ref(cat_f)
                    val_rows = parse_rows_from_ref(val_f)
                    if header_row in cat_rows and totals_row in val_rows:
                        refs_ok = True
                        return found, refs_ok  # early exit once satisfied
    return found, refs_ok


# ------------------------------ main verification ---------------------------

def verify_task(path):
    print(f"Verifying: {path}")
    score = 0.0

    # Load workbook (no points)
    try:
        wb = openpyxl.load_workbook(path, data_only=False)
        print("✓ Workbook loaded")
    except Exception as e:
        print(f"✗ Cannot load workbook: {e}")
        return 0.0

    # Choose sheet
    sheet = next(
        (s for s in wb.worksheets if s.title.lower().startswith("download")),
        wb.active,
    )
    print(f"Sheet used: {sheet.title}")

    # 1. Totals row presence --------------------------------------------------
    header_row, totals_row = locate_header_and_totals(sheet)
    if totals_row is not None:
        print(f"✓ 'Total Downloads' row found at {totals_row} (+0.35)")
        score += 0.35
    else:
        print("✗ 'Total Downloads' row missing – cannot reach full marks")
        return round(score, 2)

    # 2. Correct formulas/values ---------------------------------------------
    month_cols = month_columns(sheet, header_row or 1)
    if formulas_or_values_ok(sheet, header_row or 1, totals_row, month_cols):
        print("✓ Totals row values/formulas verified (+0.25)")
        score += 0.25
    else:
        print("✗ Totals row values/formulas incorrect")

    # 3 & 4. Line chart existence & correct data links -----------------------
    found_chart, good_refs = locate_line_chart(wb, header_row or 1, totals_row)
    if found_chart:
        print("✓ Line chart present (+0.20)")
        score += 0.20
        if good_refs:
            print("✓ Chart references correct rows (+0.20)")
            score += 0.20
        else:
            print("✗ Chart does not reference required rows")
    else:
        print("✗ No line chart detected")

    final = min(score, 1.0)
    print(f"REWARD: {final}")
    return final


if __name__ == "__main__":
    verify_task(FILE_PATH)
