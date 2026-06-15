"""
Initial Setup: Build a simple payroll calculator
Task ID: calc_wf_064
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_064'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Rates ---
    ws_rates = wb.active
    ws_rates.title = 'Rates'

    rates_headers = ['Employee ID', 'Name', 'Hourly Rate', 'Insurance Deduction', '401k %']
    for col, h in enumerate(rates_headers, 1):
        cell = ws_rates.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    rates_data = [
        ['EMP001', 'Sarah Chen', 45.00, 150.00, 0.06],
        ['EMP002', 'Marcus Johnson', 38.50, 200.00, 0.04],
        ['EMP003', 'Priya Patel', 52.00, 150.00, 0.08],
        ['EMP004', 'David Kim', 41.75, 175.00, 0.05],
        ['EMP005', 'Elena Rodriguez', 48.25, 200.00, 0.07],
        ['EMP006', 'James Wright', 35.00, 125.00, 0.03],
    ]
    for r, row_data in enumerate(rates_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_rates.cell(row=r, column=c, value=val)
            if c == 3:  # Hourly Rate
                cell.number_format = '$#,##0.00'
            elif c == 4:  # Insurance
                cell.number_format = '$#,##0.00'
            elif c == 5:  # 401k %
                cell.number_format = '0%'

    # Set column widths for Rates
    ws_rates.column_dimensions['A'].width = 14
    ws_rates.column_dimensions['B'].width = 20
    ws_rates.column_dimensions['C'].width = 14
    ws_rates.column_dimensions['D'].width = 20
    ws_rates.column_dimensions['E'].width = 10

    # --- Sheet 2: Hours ---
    ws_hours = wb.create_sheet('Hours')

    hours_headers = ['Employee ID', 'Regular Hours', 'OT Hours']
    for col, h in enumerate(hours_headers, 1):
        cell = ws_hours.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    hours_data = [
        ['EMP001', 80, 12],
        ['EMP002', 80, 5],
        ['EMP003', 76, 18],
        ['EMP004', 80, 8],
        ['EMP005', 72, 15],
        ['EMP006', 80, 3],
    ]
    for r, row_data in enumerate(hours_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_hours.cell(row=r, column=c, value=val)

    ws_hours.column_dimensions['A'].width = 14
    ws_hours.column_dimensions['B'].width = 16
    ws_hours.column_dimensions['C'].width = 12

    # --- Sheet 3: Payroll (empty, ready for calculations) ---
    ws_payroll = wb.create_sheet('Payroll')

    payroll_headers = ['Employee ID', 'Name', 'Regular Hours', 'OT Hours',
                       'Hourly Rate', 'Regular Pay', 'OT Pay', 'Gross Pay',
                       'Tax', 'Insurance', '401k', 'Total Deductions', 'Net Pay']
    for col, h in enumerate(payroll_headers, 1):
        cell = ws_payroll.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    ws_payroll.column_dimensions['A'].width = 14
    ws_payroll.column_dimensions['B'].width = 20
    ws_payroll.column_dimensions['C'].width = 15
    ws_payroll.column_dimensions['D'].width = 12
    ws_payroll.column_dimensions['E'].width = 14
    ws_payroll.column_dimensions['F'].width = 14
    ws_payroll.column_dimensions['G'].width = 12
    ws_payroll.column_dimensions['H'].width = 14
    ws_payroll.column_dimensions['I'].width = 12
    ws_payroll.column_dimensions['J'].width = 12
    ws_payroll.column_dimensions['K'].width = 12
    ws_payroll.column_dimensions['L'].width = 18
    ws_payroll.column_dimensions['M'].width = 14

    # --- Sheet 4: Summary (empty, ready for totals) ---
    ws_summary = wb.create_sheet('Summary')

    summary_headers = ['Category', 'Amount']
    for col, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 18

    # --- Sheet 5: Tax Brackets (reference) ---
    ws_tax = wb.create_sheet('Tax Brackets')

    tax_headers = ['Bracket', 'Lower Limit', 'Upper Limit', 'Rate']
    for col, h in enumerate(tax_headers, 1):
        cell = ws_tax.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    tax_data = [
        ['Bracket 1', 0, 1000, 0.10],
        ['Bracket 2', 1001, 4000, 0.22],
        ['Bracket 3', 4001, None, 0.32],
    ]
    for r, row_data in enumerate(tax_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_tax.cell(row=r, column=c, value=val)
            if c in (2, 3) and val is not None:
                cell.number_format = '$#,##0.00'
            elif c == 4:
                cell.number_format = '0%'

    # For Bracket 3, upper limit display
    ws_tax.cell(row=4, column=3, value='No limit')

    ws_tax.column_dimensions['A'].width = 14
    ws_tax.column_dimensions['B'].width = 14
    ws_tax.column_dimensions['C'].width = 14
    ws_tax.column_dimensions['D'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
