"""
Reward Script: Build a simple payroll calculator
Task ID: calc_wf_064
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Payroll sheet populated with 6 employee rows (formulas or values)
  Component 2 (0.25): Pay Stubs sheet exists with formatted stubs for all 6 employees
  Component 3 (0.25): Summary sheet has payroll totals (gross, tax, insurance, 401k, net)
  Component 4 (0.20): Payroll calculations are numerically correct (spot-check gross/tax/net)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_064'


def persist_app_state(domain):
    """Try to save any unsaved LibreOffice state."""
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print(f"PERSIST: ctrl+s sent for {domain}")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def compute_tax(gross):
    """Compute tax using the bracket system: 10% up to 1000, 22% 1001-4000, 32% above 4000."""
    if gross <= 1000:
        return gross * 0.10
    elif gross <= 4000:
        return 100 + (gross - 1000) * 0.22
    else:
        return 100 + (3000 * 0.22) + (gross - 4000) * 0.32


# Known employee data from Rates and Hours sheets (these are preconditions, NOT scored)
EMPLOYEES = {
    'EMP001': {'name': 'Sarah Chen',       'rate': 45.00,  'insurance': 150, '401k_pct': 0.06, 'reg_hrs': 80, 'ot_hrs': 12},
    'EMP002': {'name': 'Marcus Johnson',   'rate': 38.50,  'insurance': 200, '401k_pct': 0.04, 'reg_hrs': 80, 'ot_hrs': 5},
    'EMP003': {'name': 'Priya Patel',      'rate': 52.00,  'insurance': 150, '401k_pct': 0.08, 'reg_hrs': 76, 'ot_hrs': 18},
    'EMP004': {'name': 'David Kim',        'rate': 41.75,  'insurance': 175, '401k_pct': 0.05, 'reg_hrs': 80, 'ot_hrs': 8},
    'EMP005': {'name': 'Elena Rodriguez',  'rate': 48.25,  'insurance': 200, '401k_pct': 0.07, 'reg_hrs': 72, 'ot_hrs': 15},
    'EMP006': {'name': 'James Wright',     'rate': 35.00,  'insurance': 125, '401k_pct': 0.03, 'reg_hrs': 80, 'ot_hrs': 3},
}

# Precompute expected values
EXPECTED = {}
for eid, e in EMPLOYEES.items():
    reg_pay = e['reg_hrs'] * e['rate']
    ot_pay = e['ot_hrs'] * e['rate'] * 1.5
    gross = reg_pay + ot_pay
    tax = compute_tax(gross)
    k401 = gross * e['401k_pct']
    ins = e['insurance']
    total_ded = tax + ins + k401
    net = gross - total_ded
    EXPECTED[eid] = {
        'gross': gross, 'tax': tax, 'insurance': ins,
        '401k': k401, 'total_ded': total_ded, 'net': net,
    }

EXPECTED_TOTALS = {
    'gross': sum(v['gross'] for v in EXPECTED.values()),
    'tax': sum(v['tax'] for v in EXPECTED.values()),
    'insurance': sum(v['insurance'] for v in EXPECTED.values()),
    '401k': sum(v['401k'] for v in EXPECTED.values()),
    'total_ded': sum(v['total_ded'] for v in EXPECTED.values()),
    'net': sum(v['net'] for v in EXPECTED.values()),
}


def verify_task(file_path):
    """Verify task completion with progressive scoring. Returns float 0.0-1.0."""
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Also load with data_only to get computed values
    try:
        wb_data = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"WARN: Cannot load data_only: {e}")
        wb_data = None

    # =========================================================================
    # Component 1: Payroll sheet has data rows for 6 employees (0.30 points)
    # Initial has only headers (row 1), golden has rows 2-7 with formulas
    # =========================================================================
    try:
        if 'Payroll' not in wb.sheetnames:
            print("FAIL: Component 1 - Payroll sheet not found")
        else:
            ws_payroll = wb['Payroll']
            # Count non-empty data rows (rows 2+)
            emp_count = 0
            has_formulas_or_values = 0
            for r in range(2, ws_payroll.max_row + 1):
                emp_id = ws_payroll.cell(row=r, column=1).value
                if emp_id is not None:
                    emp_count += 1
                    # Check that at least some calculation columns are populated
                    # (F=Regular Pay, H=Gross Pay, M=Net Pay)
                    h_val = ws_payroll.cell(row=r, column=8).value  # Gross Pay
                    m_val = ws_payroll.cell(row=r, column=13).value  # Net Pay
                    if h_val is not None and m_val is not None:
                        has_formulas_or_values += 1

            if emp_count >= 6 and has_formulas_or_values >= 6:
                print(f"PASS: Component 1 - Payroll has {emp_count} employees with calculations (0.30 pts)")
                total_score += 0.30
            elif emp_count >= 4 and has_formulas_or_values >= 4:
                print(f"PARTIAL: Component 1 - Payroll has {emp_count} employees, {has_formulas_or_values} with calcs (0.15 pts)")
                total_score += 0.15
            elif emp_count >= 1:
                print(f"PARTIAL: Component 1 - Payroll has {emp_count} employees, {has_formulas_or_values} with calcs (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 1 - Payroll has no employee data rows")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # =========================================================================
    # Component 2: Pay Stubs sheet exists with stubs for all 6 employees (0.25 points)
    # This sheet does NOT exist in the initial file
    # =========================================================================
    try:
        pay_stubs_sheet = None
        for sn in wb.sheetnames:
            if 'pay' in sn.lower() and 'stub' in sn.lower():
                pay_stubs_sheet = sn
                break

        if pay_stubs_sheet is None:
            print("FAIL: Component 2 - No Pay Stubs sheet found")
        else:
            ws_stubs = wb[pay_stubs_sheet]
            # Count how many employee stubs exist by looking for "PAY STUB" headers
            # or employee name references
            stub_count = 0
            employee_names_found = set()
            for r in range(1, ws_stubs.max_row + 1):
                cell_a = ws_stubs.cell(row=r, column=1).value
                cell_b = ws_stubs.cell(row=r, column=2).value
                if cell_a and 'PAY STUB' in str(cell_a).upper():
                    stub_count += 1
                # Check for employee names in column B next to "Employee:" label
                if cell_a and 'employee' in str(cell_a).lower() and cell_b:
                    for eid, emp in EMPLOYEES.items():
                        if emp['name'] in str(cell_b):
                            employee_names_found.add(emp['name'])

            if stub_count >= 6 and len(employee_names_found) >= 6:
                print(f"PASS: Component 2 - Pay Stubs has {stub_count} stubs for {len(employee_names_found)} employees (0.25 pts)")
                total_score += 0.25
            elif stub_count >= 4:
                print(f"PARTIAL: Component 2 - Pay Stubs has {stub_count} stubs, {len(employee_names_found)} names found (0.15 pts)")
                total_score += 0.15
            elif stub_count >= 1:
                print(f"PARTIAL: Component 2 - Pay Stubs has {stub_count} stubs (0.08 pts)")
                total_score += 0.08
            else:
                print(f"FAIL: Component 2 - Pay Stubs sheet exists but no stubs found")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # =========================================================================
    # Component 3: Summary sheet has payroll totals (0.25 points)
    # Initial summary has only headers, golden has computed totals
    # =========================================================================
    try:
        if 'Summary' not in wb.sheetnames:
            print("FAIL: Component 3 - Summary sheet not found")
        else:
            ws_summary = wb['Summary']
            # Check that summary has at least gross/tax/net totals populated
            summary_values = {}
            for r in range(1, ws_summary.max_row + 1):
                label = ws_summary.cell(row=r, column=1).value
                value = ws_summary.cell(row=r, column=2).value
                if label and value is not None:
                    label_lower = str(label).lower()
                    summary_values[label_lower] = value

            # We need at least: total gross, total tax, total net
            has_gross = any('gross' in k for k in summary_values)
            has_tax = any('tax' in k for k in summary_values)
            has_net = any('net' in k for k in summary_values)
            has_deductions = any('deduction' in k for k in summary_values)
            has_insurance = any('insurance' in k for k in summary_values)
            has_401k = any('401k' in k or '401' in k for k in summary_values)

            checks_passed = sum([has_gross, has_tax, has_net, has_deductions])

            if checks_passed >= 4:
                print(f"PASS: Component 3 - Summary has gross/tax/deductions/net totals (0.25 pts)")
                total_score += 0.25
            elif checks_passed >= 2:
                print(f"PARTIAL: Component 3 - Summary has {checks_passed}/4 key totals (0.12 pts)")
                total_score += 0.12
            elif checks_passed >= 1:
                print(f"PARTIAL: Component 3 - Summary has {checks_passed}/4 key totals (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 3 - Summary has no payroll totals (only headers)")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # =========================================================================
    # Component 4: Payroll calculations are numerically correct (0.20 points)
    # Spot-check computed values against expected (using data_only or formulas)
    # =========================================================================
    try:
        if wb_data is None or 'Payroll' not in wb_data.sheetnames:
            # Fallback: check formula patterns in the regular workbook
            ws_p = wb['Payroll']
            formula_checks_passed = 0
            for r in range(2, min(ws_p.max_row + 1, 8)):
                # Check gross pay formula contains addition of regular and OT
                h_val = ws_p.cell(row=r, column=8).value
                i_val = ws_p.cell(row=r, column=9).value  # Tax
                m_val = ws_p.cell(row=r, column=13).value  # Net Pay
                if h_val and isinstance(h_val, str) and ('F' in h_val or 'G' in h_val or '+' in h_val):
                    formula_checks_passed += 1
                if i_val and isinstance(i_val, str) and 'IF' in str(i_val).upper():
                    formula_checks_passed += 1
                if m_val and isinstance(m_val, str):
                    formula_checks_passed += 1

            if formula_checks_passed >= 12:
                print(f"PASS: Component 4 - {formula_checks_passed} formula checks passed (0.20 pts)")
                total_score += 0.20
            elif formula_checks_passed >= 6:
                print(f"PARTIAL: Component 4 - {formula_checks_passed}/18 formula checks passed (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 4 - Only {formula_checks_passed} formula checks passed")
        else:
            ws_data = wb_data['Payroll']
            # Spot check: look at cached computed values for correctness
            correct_values = 0
            total_checks = 0

            for r in range(2, min(ws_data.max_row + 1, 8)):
                emp_id_cell = ws_data.cell(row=r, column=1).value
                if emp_id_cell is None:
                    continue

                emp_id = str(emp_id_cell).strip()
                if emp_id not in EXPECTED:
                    continue

                exp = EXPECTED[emp_id]
                tol = 1.0  # tolerance for rounding

                # Check Gross Pay (col H=8)
                gross_val = ws_data.cell(row=r, column=8).value
                total_checks += 1
                if gross_val is not None:
                    try:
                        if abs(float(gross_val) - exp['gross']) < tol:
                            correct_values += 1
                    except (ValueError, TypeError):
                        pass

                # Check Tax (col I=9)
                tax_val = ws_data.cell(row=r, column=9).value
                total_checks += 1
                if tax_val is not None:
                    try:
                        if abs(float(tax_val) - exp['tax']) < tol:
                            correct_values += 1
                    except (ValueError, TypeError):
                        pass

                # Check Net Pay (col M=13)
                net_val = ws_data.cell(row=r, column=13).value
                total_checks += 1
                if net_val is not None:
                    try:
                        if abs(float(net_val) - exp['net']) < tol:
                            correct_values += 1
                    except (ValueError, TypeError):
                        pass

            if total_checks == 0:
                # No data_only values available, fall back to formula check
                ws_p = wb['Payroll']
                formula_count = 0
                for r in range(2, min(ws_p.max_row + 1, 8)):
                    for c in [6, 7, 8, 9, 13]:  # F, G, H, I, M
                        val = ws_p.cell(row=r, column=c).value
                        if val and isinstance(val, str) and val.startswith('='):
                            formula_count += 1
                if formula_count >= 20:
                    print(f"PASS: Component 4 - {formula_count} formulas present in Payroll (0.20 pts)")
                    total_score += 0.20
                elif formula_count >= 10:
                    print(f"PARTIAL: Component 4 - {formula_count} formulas in Payroll (0.10 pts)")
                    total_score += 0.10
                else:
                    print(f"FAIL: Component 4 - Only {formula_count} formulas in Payroll")
            else:
                ratio = correct_values / total_checks if total_checks > 0 else 0
                if ratio >= 0.8:
                    print(f"PASS: Component 4 - {correct_values}/{total_checks} values correct (0.20 pts)")
                    total_score += 0.20
                elif ratio >= 0.5:
                    print(f"PARTIAL: Component 4 - {correct_values}/{total_checks} values correct (0.10 pts)")
                    total_score += 0.10
                else:
                    print(f"FAIL: Component 4 - Only {correct_values}/{total_checks} values correct")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
