"""
Reward Script: Sales pipeline aging report with color-coded deal health indicators
Task ID: calc_sales_055
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): E2:E7 contain =TODAY()-D<row> formulas for days calculation
  Component 2 (0.30): F2:F7 contain nested IF formulas for health classification
  Component 3 (0.40): Conditional formatting on F column with green/yellow/red colors
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_055'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Aging' sheet must exist
    if 'Aging' not in wb.sheetnames:
        print("CRITICAL: 'Aging' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Aging']

    # Component 1: E2:E7 contain =TODAY()-D<row> formulas (0.30 points)
    # These cells are EMPTY in the initial env, so this only passes on golden
    try:
        e_formula_count = 0
        for row_num in range(2, 8):
            cell_val = ws.cell(row=row_num, column=5).value  # column E
            if cell_val is not None and isinstance(cell_val, str):
                # Normalize: remove spaces, uppercase
                normalized = cell_val.upper().replace(" ", "")
                expected = f"=TODAY()-D{row_num}"
                if normalized == expected:
                    e_formula_count += 1
                else:
                    print(f"  INFO: E{row_num} has formula '{cell_val}', expected '{expected}'")
            else:
                print(f"  INFO: E{row_num} value is {repr(cell_val)} (not a formula string)")

        if e_formula_count == 6:
            print(f"PASS: Component 1 -- All 6 E-column formulas correct (0.30 pts)")
            total_score += 0.30
        elif e_formula_count >= 3:
            partial = round(0.30 * (e_formula_count / 6), 2)
            print(f"PARTIAL: Component 1 -- {e_formula_count}/6 E-column formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 -- Only {e_formula_count}/6 E-column formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: F2:F7 contain nested IF formulas for health classification (0.30 points)
    # These cells are EMPTY in the initial env, so this only passes on golden
    try:
        f_formula_count = 0
        for row_num in range(2, 8):
            cell_val = ws.cell(row=row_num, column=6).value  # column F
            if cell_val is not None and isinstance(cell_val, str):
                normalized = cell_val.upper().replace(" ", "").replace("'", '"')
                # Expected pattern: =IF(E<row>>60,"CRITICAL",IF(E<row>>30,"ATRISK","HEALTHY"))
                # Allow flexible quoting and spacing
                expected_pattern = f'=IF(E{row_num}>60,"CRITICAL",IF(E{row_num}>30,"ATRISK","HEALTHY"))'
                if normalized == expected_pattern:
                    f_formula_count += 1
                else:
                    # Also try >=60 and >=30 variants
                    alt1 = f'=IF(E{row_num}>=60,"CRITICAL",IF(E{row_num}>=30,"ATRISK","HEALTHY"))'
                    if normalized == alt1:
                        f_formula_count += 1
                    else:
                        print(f"  INFO: F{row_num} has formula '{cell_val}' (normalized: {normalized})")
            else:
                print(f"  INFO: F{row_num} value is {repr(cell_val)} (not a formula string)")

        if f_formula_count == 6:
            print(f"PASS: Component 2 -- All 6 F-column IF formulas correct (0.30 pts)")
            total_score += 0.30
        elif f_formula_count >= 3:
            partial = round(0.30 * (f_formula_count / 6), 2)
            print(f"PARTIAL: Component 2 -- {f_formula_count}/6 F-column IF formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- Only {f_formula_count}/6 F-column IF formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Conditional formatting on F2:F7 with correct colors (0.40 points)
    # No conditional formatting exists in the initial env
    try:
        cf_rules_found = []
        target_range_found = False

        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            # Check if the conditional formatting covers the F column data range
            if 'F' in cf_range:
                target_range_found = True
                for rule in cf.rules:
                    if rule.dxf and rule.dxf.fill:
                        try:
                            fg_rgb = rule.dxf.fill.fgColor.rgb if rule.dxf.fill.fgColor else None
                        except:
                            fg_rgb = None

                        formula_str = str(rule.formula) if hasattr(rule, 'formula') and rule.formula else ''

                        cf_rules_found.append({
                            'type': rule.type,
                            'operator': getattr(rule, 'operator', None),
                            'formula': formula_str,
                            'fill_rgb': fg_rgb,
                        })

        # Check for the three required color rules
        has_green = False  # Healthy
        has_yellow = False  # At Risk
        has_red = False  # Critical

        # Known color mappings from the golden file:
        # Green (Healthy): FF00B050
        # Yellow/Orange (At Risk): FFFFC000
        # Red (Critical): FFFF0000
        green_colors = {'FF00B050', 'FF00FF00', 'FF92D050', 'FF00B050'}
        yellow_colors = {'FFFFC000', 'FFFFFF00', 'FFFFD700', 'FFFFC000'}
        red_colors = {'FFFF0000', 'FFFF4444', 'FFCC0000', 'FFFF0000'}

        for rule_info in cf_rules_found:
            fill_rgb = rule_info.get('fill_rgb', '')
            formula = rule_info.get('formula', '').upper()
            if fill_rgb in green_colors or ('HEALTHY' in formula and fill_rgb):
                has_green = True
            if fill_rgb in yellow_colors or ('ATRISK' in formula and fill_rgb) or ('AT RISK' in formula.replace('"', '') and fill_rgb):
                has_yellow = True
            if fill_rgb in red_colors or ('CRITICAL' in formula and fill_rgb):
                has_red = True

        color_count = sum([has_green, has_yellow, has_red])

        if target_range_found and color_count == 3:
            print(f"PASS: Component 3 -- Conditional formatting with all 3 health colors (0.40 pts)")
            total_score += 0.40
        elif target_range_found and color_count >= 1:
            partial = round(0.40 * (color_count / 3), 2)
            print(f"PARTIAL: Component 3 -- {color_count}/3 conditional formatting colors found ({partial} pts)")
            total_score += partial
        else:
            if not target_range_found:
                print(f"FAIL: Component 3 -- No conditional formatting found on F column")
            else:
                print(f"FAIL: Component 3 -- Conditional formatting found but no health color rules detected")
            print(f"  DEBUG: Found {len(cf_rules_found)} CF rules: {cf_rules_found}")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
