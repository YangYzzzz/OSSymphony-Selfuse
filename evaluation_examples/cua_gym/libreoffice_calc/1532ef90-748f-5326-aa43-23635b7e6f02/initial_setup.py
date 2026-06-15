"""
Initial Setup: Sales pipeline aging report with deal data but no formulas or formatting.
Task ID: calc_sales_055
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_055'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Aging ---
    ws = wb.active
    ws.title = 'Aging'

    # Headers
    headers = ['Deal', 'Value', 'Stage', 'Stage Entry Date', 'Days in Stage', 'Health']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    white_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows - columns A through D only
    # E (Days in Stage) and F (Health) are LEFT EMPTY for the agent to fill
    data = [
        ['Deal A', 150000, 'Proposal',    '2026-03-15'],
        ['Deal B', 80000,  'Negotiation',  '2026-02-01'],
        ['Deal C', 200000, 'Discovery',    '2026-01-10'],
        ['Deal D', 95000,  'Proposal',     '2026-03-25'],
        ['Deal E', 300000, 'Negotiation',  '2026-03-01'],
        ['Deal F', 120000, 'Discovery',    '2025-12-15'],
    ]

    currency_format = '$#,##0'
    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])  # Deal name
        val_cell = ws.cell(row=r, column=2, value=row_data[1])  # Value
        val_cell.number_format = currency_format
        ws.cell(row=r, column=3, value=row_data[2])  # Stage
        date_cell = ws.cell(row=r, column=4, value=row_data[3])  # Stage Entry Date
        # Columns E and F intentionally left empty

    # Column widths for readability
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
