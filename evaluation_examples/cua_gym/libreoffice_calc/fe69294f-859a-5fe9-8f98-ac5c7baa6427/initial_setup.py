"""
Initial Setup: ArXiv cs.CL papers tracker - weekly_arxiv.ods
Task ID: osworld_multi_apps_arxiv_llms_calc_012
Domain: libreoffice_calc (multi-app: Chrome + LibreOffice Calc)

Creates weekly_arxiv.ods with:
  - Sheet1: Headers only (arXiv ID, Title, Authors, Abstract, Keywords)
  - Keyword Frequency sheet: Headers only (Keyword, Count)
  - Chrome open to arxiv.org/list/cs.CL/2024-02
  - LibreOffice Calc open with weekly_arxiv.ods
"""

import os
import shlex
import subprocess
import time
import tempfile

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_arxiv_llms_calc_012'
OUTPUT = f'{WORKDIR}/weekly_arxiv.ods'
TEMP_XLSX = f'{WORKDIR}/weekly_arxiv_temp.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # --- Sheet 1: Paper data ---
    ws1 = wb.active
    ws1.title = 'Sheet1'

    # Headers for paper data
    headers = ['arXiv ID', 'Title', 'Authors', 'Abstract', 'Keywords']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFD3E4F7", end_color="FFD3E4F7", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=False)

    # Set column widths
    ws1.column_dimensions['A'].width = 20  # arXiv ID
    ws1.column_dimensions['B'].width = 50  # Title
    ws1.column_dimensions['C'].width = 40  # Authors
    ws1.column_dimensions['D'].width = 80  # Abstract
    ws1.column_dimensions['E'].width = 40  # Keywords

    # NO DATA ROWS — task requires agent to fill them in

    # --- Sheet 2: Keyword Frequency ---
    ws2 = wb.create_sheet('Keyword Frequency')

    # Headers for keyword frequency
    kw_headers = ['Keyword', 'Count']
    for col, h in enumerate(kw_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFD3E4F7", end_color="FFD3E4F7", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Set column widths
    ws2.column_dimensions['A'].width = 30  # Keyword
    ws2.column_dimensions['B'].width = 15  # Count

    # NO DATA ROWS — task requires agent to fill them in

    # Save as temp xlsx first
    wb.save(TEMP_XLSX)
    print(f'Temp xlsx created: {TEMP_XLSX}')

    # Convert xlsx to ods using LibreOffice headless
    env = os.environ.copy()
    env['DISPLAY'] = ':0'
    result = subprocess.run(
        ['libreoffice', '--headless', '--convert-to', 'ods', '--outdir', WORKDIR, TEMP_XLSX],
        capture_output=True,
        text=True,
        env=env,
        timeout=60
    )
    print(f'LibreOffice convert stdout: {result.stdout}')
    print(f'LibreOffice convert stderr: {result.stderr}')

    # The converted file will be named weekly_arxiv_temp.ods, rename to weekly_arxiv.ods
    converted = f'{WORKDIR}/weekly_arxiv_temp.ods'
    if os.path.exists(converted):
        if os.path.exists(OUTPUT):
            os.remove(OUTPUT)
        os.rename(converted, OUTPUT)
        print(f'Renamed {converted} to {OUTPUT}')
    elif os.path.exists(OUTPUT):
        print(f'Output already exists: {OUTPUT}')
    else:
        # Fallback: try alternative conversion approach
        print(f'WARNING: Conversion may have failed. Checking for output...')
        import glob
        ods_files = glob.glob(f'{WORKDIR}/*.ods')
        print(f'Found .ods files: {ods_files}')

    # Remove temp xlsx
    if os.path.exists(TEMP_XLSX):
        os.remove(TEMP_XLSX)
        print(f'Removed temp file: {TEMP_XLSX}')

    if os.path.exists(OUTPUT):
        print(f'Initial file created successfully: {OUTPUT}')
    else:
        print(f'ERROR: Output file not found at {OUTPUT}')

    # GUI-ready startup: open Chrome with arXiv cs.CL Feb 2024 page first
    launch_gui(
        'google-chrome "https://arxiv.org/list/cs.CL/2024-02"',
        delay_sec=3.0
    )
    print('Launched Chrome with arXiv cs.CL Feb 2024')

    # Open LibreOffice Calc with the ods file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=3.0)
    print('Launched LibreOffice Calc with weekly_arxiv.ods')

    print('GUI_READY: launched Chrome and LibreOffice Calc with DISPLAY=:0')


create_initial()
