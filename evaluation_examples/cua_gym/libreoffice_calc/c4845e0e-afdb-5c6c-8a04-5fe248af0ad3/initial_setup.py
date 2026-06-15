"""
Initial Setup: Build a mini-dashboard on CompanyView referencing Sales and Expenses sheets.
Task ID: calc_mcp_065
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_065'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Sales ---
    ws_sales = wb.active
    ws_sales.title = 'Sales'

    # Headers
    sales_headers = ['Month', 'Region', 'Product', 'Revenue']
    for col, h in enumerate(sales_headers, 1):
        cell = ws_sales.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Generate 48 rows of realistic sales data (rows 2-49)
    regions = ['North', 'South', 'East', 'West']
    products = ['Widget Pro', 'Widget Lite', 'Widget Ultra', 'Widget Basic']
    months = [
        'Jan 2025', 'Feb 2025', 'Mar 2025', 'Apr 2025',
        'May 2025', 'Jun 2025', 'Jul 2025', 'Aug 2025',
        'Sep 2025', 'Oct 2025', 'Nov 2025', 'Dec 2025',
    ]

    # Pre-calculated revenue values that sum to 1500000 in D50
    # We'll fill rows 2-49 with values, then put the total in row 50
    revenue_values = [
        28500, 31200, 34800, 27600, 29900, 33100, 35400, 26800,
        30700, 32500, 28900, 31800, 34200, 27100, 29500, 33600,
        35800, 26300, 30200, 32900, 28400, 31500, 34500, 27800,
        30100, 33300, 35100, 26500, 29700, 32200, 28700, 31900,
        34400, 27300, 29800, 33400, 35600, 26100, 30400, 32700,
        28600, 31600, 34600, 27500, 30000, 33200, 35300, 26700,
    ]
    # Adjust last value so total = 1500000
    current_sum = sum(revenue_values)
    # We need D50 = 1500000 as a total. Rows 2-49 are data, row 50 is total.
    # The total in D50 should be exactly 1500000
    # Let's adjust the last value
    revenue_values[-1] = revenue_values[-1] + (1500000 - current_sum)

    for i, rev in enumerate(revenue_values):
        row = i + 2
        month_idx = i % 12
        region_idx = i % 4
        product_idx = (i // 4) % 4
        ws_sales.cell(row=row, column=1, value=months[month_idx])
        ws_sales.cell(row=row, column=2, value=regions[region_idx])
        ws_sales.cell(row=row, column=3, value=products[product_idx])
        ws_sales.cell(row=row, column=4, value=rev)

    # Row 50: Total row
    ws_sales.cell(row=50, column=1, value='TOTAL')
    ws_sales.cell(row=50, column=1).font = Font(bold=True)
    ws_sales.cell(row=50, column=4, value=1500000)
    ws_sales.cell(row=50, column=4).font = Font(bold=True)
    ws_sales.cell(row=50, column=4).number_format = '#,##0'

    # Set column widths for readability
    ws_sales.column_dimensions['A'].width = 14
    ws_sales.column_dimensions['B'].width = 12
    ws_sales.column_dimensions['C'].width = 16
    ws_sales.column_dimensions['D'].width = 14

    # --- Sheet 2: Expenses ---
    ws_exp = wb.create_sheet('Expenses')

    exp_headers = ['Month', 'Department', 'Category', 'Amount']
    for col, h in enumerate(exp_headers, 1):
        cell = ws_exp.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    departments = ['Engineering', 'Marketing', 'Operations', 'HR']
    categories = ['Salaries', 'Software', 'Travel', 'Office Supplies']

    expense_values = [
        18500, 21200, 19800, 17600, 20900, 22100, 19400, 18800,
        21700, 20500, 18900, 22800, 19200, 17100, 20500, 21600,
        19800, 18300, 21200, 20900, 18400, 22500, 19500, 17800,
        21100, 20300, 19100, 18500, 20700, 22200, 19600, 18900,
        21400, 20600, 18800, 22400, 19300, 17200, 20800, 21700,
        19900, 18600, 21300, 20800, 18500, 22300, 19400, 17700,
    ]
    current_sum_exp = sum(expense_values)
    expense_values[-1] = expense_values[-1] + (980000 - current_sum_exp)

    for i, amt in enumerate(expense_values):
        row = i + 2
        month_idx = i % 12
        dept_idx = i % 4
        cat_idx = (i // 4) % 4
        ws_exp.cell(row=row, column=1, value=months[month_idx])
        ws_exp.cell(row=row, column=2, value=departments[dept_idx])
        ws_exp.cell(row=row, column=3, value=categories[cat_idx])
        ws_exp.cell(row=row, column=4, value=amt)

    # Row 50: Total row
    ws_exp.cell(row=50, column=1, value='TOTAL')
    ws_exp.cell(row=50, column=1).font = Font(bold=True)
    ws_exp.cell(row=50, column=4, value=980000)
    ws_exp.cell(row=50, column=4).font = Font(bold=True)
    ws_exp.cell(row=50, column=4).number_format = '#,##0'

    ws_exp.column_dimensions['A'].width = 14
    ws_exp.column_dimensions['B'].width = 16
    ws_exp.column_dimensions['C'].width = 18
    ws_exp.column_dimensions['D'].width = 14

    # --- Sheet 3: CompanyView (dashboard shell) ---
    ws_cv = wb.create_sheet('CompanyView')

    # Title
    ws_cv.cell(row=1, column=1, value='Company Dashboard')
    ws_cv.cell(row=1, column=1).font = Font(bold=True, size=14)

    # Labels in A2:A5
    ws_cv.cell(row=2, column=1, value='Revenue')
    ws_cv.cell(row=3, column=1, value='Costs')
    ws_cv.cell(row=4, column=1, value='Profit')
    ws_cv.cell(row=5, column=1, value='Margin')

    # B2:B5 intentionally left EMPTY (task requires agent to fill these)

    # Style labels
    for r in range(2, 6):
        ws_cv.cell(row=r, column=1).font = Font(bold=True)

    ws_cv.column_dimensions['A'].width = 16
    ws_cv.column_dimensions['B'].width = 18

    # Reorder: CompanyView first for convenience (but task doesn't require it)
    # Keep Sales first as that's more natural for a workbook

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
