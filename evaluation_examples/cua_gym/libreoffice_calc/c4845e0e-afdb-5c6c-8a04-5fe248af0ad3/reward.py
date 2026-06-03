"""
Reward Script: Mini-dashboard on CompanyView sheet
Task ID: calc_mcp_065
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): B2 references Sales!D50 (revenue)
  Component 2 (0.25): B3 references Expenses!D50 (costs)
  Component 3 (0.25): B4 calculates profit as =B2-B3
  Component 4 (0.25): B5 calculates margin as =B4/B2 with percentage format
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_065'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def check_formula(ws, coord, expected_patterns):
    """
    Check if a cell contains a formula matching any of the expected patterns.
    Patterns are compared case-insensitively with spaces removed.
    Returns (True, actual_value) or (False, actual_value).
    """
    val = ws[coord].value
    if val is None:
        return False, None
    if not isinstance(val, str) or not val.startswith('='):
        return False, val
    normalized = val.upper().replace(" ", "")
    for pattern in expected_patterns:
        if normalized == pattern.upper().replace(" ", ""):
            return True, val
    return False, val


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that CompanyView sheet exists
    if 'CompanyView' not in wb.sheetnames:
        print("CRITICAL: 'CompanyView' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['CompanyView']

    # Component 1: B2 references Sales!D50 for revenue (0.25 points)
    # LibreOffice uses dot notation (Sales.D50), Excel uses bang (Sales!D50)
    # Accept both variants
    try:
        expected_b2 = ["=Sales.D50", "=Sales!D50", "='Sales'.D50", "='Sales'!D50"]
        passed, actual = check_formula(ws, 'B2', expected_b2)
        if passed:
            print(f"PASS: Component 1 — B2 references Sales D50: {actual} (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — B2 expected reference to Sales!D50, found: {actual}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: B3 references Expenses!D50 for costs (0.25 points)
    try:
        expected_b3 = ["=Expenses.D50", "=Expenses!D50", "='Expenses'.D50", "='Expenses'!D50"]
        passed, actual = check_formula(ws, 'B3', expected_b3)
        if passed:
            print(f"PASS: Component 2 — B3 references Expenses D50: {actual} (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — B3 expected reference to Expenses!D50, found: {actual}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: B4 calculates profit as =B2-B3 (0.25 points)
    try:
        expected_b4 = ["=B2-B3"]
        passed, actual = check_formula(ws, 'B4', expected_b4)
        if passed:
            print(f"PASS: Component 3 — B4 profit formula: {actual} (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 3 — B4 expected =B2-B3, found: {actual}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: B5 calculates margin as =B4/B2 with percentage format (0.25 points)
    try:
        expected_b5 = ["=B4/B2"]
        passed, actual = check_formula(ws, 'B5', expected_b5)
        if passed:
            # Also check percentage format
            fmt = ws['B5'].number_format
            is_pct = '%' in str(fmt)
            if is_pct:
                print(f"PASS: Component 4 — B5 margin formula: {actual}, format: {fmt} (0.25 pts)")
                total_score += 0.25
            elif not is_pct:
                # Partial: formula correct but format not percentage
                print(f"PARTIAL: Component 4 — B5 formula correct ({actual}) but format is '{fmt}', not percentage (0.15 pts)")
                total_score += 0.15
        else:
            print(f"FAIL: Component 4 — B5 expected =B4/B2, found: {actual}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
