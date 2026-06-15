"""
Initial Setup: Sales lead list with duplicates and blank emails for cleanup task
Task ID: calc_sales_pipeline_duplicate_cleanup_049
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_pipeline_duplicate_cleanup_049'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'LeadList'

    # Headers
    headers = ['Lead ID', 'First Name', 'Last Name', 'Company', 'Email', 'Source', 'Date Added', 'Duplicate']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFFFF')

    # Realistic data pools
    first_names = [
        'James', 'Maria', 'Robert', 'Linda', 'Michael', 'Barbara', 'William', 'Patricia',
        'David', 'Jennifer', 'Richard', 'Margaret', 'Joseph', 'Susan', 'Thomas', 'Dorothy',
        'Charles', 'Lisa', 'Christopher', 'Nancy', 'Daniel', 'Karen', 'Matthew', 'Betty',
        'Anthony', 'Helen', 'Mark', 'Sandra', 'Donald', 'Donna', 'Steven', 'Carol',
        'Paul', 'Ruth', 'Andrew', 'Sharon', 'Joshua', 'Michelle', 'Kenneth', 'Laura',
        'Kevin', 'Sarah', 'Brian', 'Kimberly', 'George', 'Deborah', 'Timothy', 'Jessica',
        'Ronald', 'Shirley', 'Edward', 'Cynthia', 'Jason', 'Angela', 'Jeffrey', 'Melissa',
        'Ryan', 'Brenda', 'Jacob', 'Amy', 'Gary', 'Anna', 'Nicholas', 'Rebecca',
        'Eric', 'Virginia', 'Jonathan', 'Kathleen', 'Stephen', 'Pamela', 'Larry', 'Martha',
        'Justin', 'Debra', 'Scott', 'Amanda', 'Brandon', 'Stephanie', 'Raymond', 'Carolyn'
    ]

    last_names = [
        'Smith', 'Johnson', 'Williams', 'Brown', 'Jones', 'Garcia', 'Miller', 'Davis',
        'Rodriguez', 'Martinez', 'Hernandez', 'Lopez', 'Gonzalez', 'Wilson', 'Anderson',
        'Thomas', 'Taylor', 'Moore', 'Jackson', 'Martin', 'Lee', 'Perez', 'Thompson',
        'White', 'Harris', 'Sanchez', 'Clark', 'Ramirez', 'Lewis', 'Robinson', 'Walker',
        'Young', 'Allen', 'King', 'Wright', 'Scott', 'Torres', 'Nguyen', 'Hill', 'Flores',
        'Green', 'Adams', 'Nelson', 'Baker', 'Hall', 'Rivera', 'Campbell', 'Mitchell',
        'Carter', 'Roberts', 'Turner', 'Phillips', 'Evans', 'Collins', 'Stewart', 'Morris',
        'Rogers', 'Reed', 'Cook', 'Morgan', 'Bell', 'Murphy', 'Bailey', 'Cooper',
        'Richardson', 'Cox', 'Howard', 'Ward', 'Torres', 'Peterson', 'Gray', 'Ramirez',
        'James', 'Watson', 'Brooks', 'Kelly', 'Sanders', 'Price', 'Bennett', 'Wood'
    ]

    companies = [
        'Acme Corp', 'TechNova Solutions', 'BlueSky Enterprises', 'Pinnacle Systems',
        'Meridian Group', 'Apex Technologies', 'Horizon Digital', 'NextGen Analytics',
        'GlobalCore Inc', 'Stellar Dynamics', 'Catalyst Ventures', 'Paradigm Shift LLC',
        'Velocity Partners', 'Synapse Technologies', 'Quantum Leap Corp', 'BrightPath Consulting',
        'Infinity Loop Systems', 'Phoenix Rising LLC', 'Summit Digital Works', 'Crossroads Solutions',
        'Harbor Bay Systems', 'Ironwood Analytics', 'Clearview Technologies', 'Keystone Ventures',
        'Alpine Solutions Group', 'Riverview Consulting', 'Cornerstone Digital', 'Trailhead Corp',
        'Vanguard Systems', 'Lakeside Analytics', 'Orion Technologies', 'Redwood Enterprises',
        'Eclipse Solutions', 'Gateway Digital', 'Frontier Systems', 'Cascade Technologies',
        'Bedrock Solutions', 'Skyline Enterprises', 'Northgate Systems', 'Westbrook Consulting',
        'Eastpoint Solutions', 'Southridge Technologies', 'Crestview Digital', 'Briarwood Corp',
        'Meadowbrook Systems', 'Rockford Analytics', 'Springfield Digital', 'Greenfield Solutions',
        'Oakdale Technologies', 'Elmwood Enterprises'
    ]

    domains = [
        'gmail.com', 'yahoo.com', 'outlook.com', 'hotmail.com', 'company.com',
        'techcorp.net', 'businessmail.com', 'enterprise.org', 'protonmail.com', 'icloud.com'
    ]

    sources = ['Website', 'LinkedIn', 'Trade Show', 'Referral', 'Cold Outreach', 'Webinar', 'Email Campaign']

    months = ['2024-01', '2024-02', '2024-03', '2024-04', '2024-05', '2024-06',
              '2024-07', '2024-08', '2024-09', '2024-10', '2024-11', '2024-12',
              '2025-01', '2025-02', '2025-03']

    def make_date():
        month = random.choice(months)
        day = random.randint(1, 28)
        return f'{month}-{day:02d}'

    # Build a pool of unique leads (365 unique entries)
    leads = []
    used_combos = set()
    lead_id = 1001

    while len(leads) < 365:
        fn = random.choice(first_names)
        ln = random.choice(last_names)
        company = random.choice(companies)
        email_user = f'{fn.lower()}.{ln.lower()}{random.randint(1, 99)}'
        email_domain = random.choice(domains)
        email = f'{email_user}@{email_domain}'
        combo = (company, email)
        if combo not in used_combos:
            used_combos.add(combo)
            leads.append({
                'id': lead_id,
                'fn': fn,
                'ln': ln,
                'company': company,
                'email': email,
                'source': random.choice(sources),
                'date': make_date()
            })
            lead_id += 1

    # Pick ~30 leads to be duplicated (appear 2x or 3x)
    duplicate_pool = random.sample(leads[:200], 30)
    extra_rows = []
    for lead in duplicate_pool:
        times = random.choice([2, 3])  # duplicate 2 or 3 times total
        for _ in range(times - 1):
            extra_rows.append({
                'id': lead_id,
                'fn': lead['fn'],
                'ln': lead['ln'],
                'company': lead['company'],
                'email': lead['email'],
                'source': random.choice(sources),  # may differ (imported from different source)
                'date': make_date()
            })
            lead_id += 1

    # Add some leads with blank emails (~15 blank email rows)
    blank_email_leads = []
    for _ in range(15):
        fn = random.choice(first_names)
        ln = random.choice(last_names)
        company = random.choice(companies)
        blank_email_leads.append({
            'id': lead_id,
            'fn': fn,
            'ln': ln,
            'company': company,
            'email': '',   # blank email
            'source': random.choice(sources),
            'date': make_date()
        })
        lead_id += 1

    # Combine: unique leads + duplicates + blank emails, shuffle, then trim to 500
    all_rows = leads + extra_rows + blank_email_leads
    random.shuffle(all_rows)
    all_rows = all_rows[:500]

    # Write data rows
    for r, row_data in enumerate(all_rows, 2):
        ws.cell(row=r, column=1, value=row_data['id'])
        ws.cell(row=r, column=2, value=row_data['fn'])
        ws.cell(row=r, column=3, value=row_data['ln'])
        ws.cell(row=r, column=4, value=row_data['company'])
        ws.cell(row=r, column=5, value=row_data['email'] if row_data['email'] else None)
        ws.cell(row=r, column=6, value=row_data['source'])
        ws.cell(row=r, column=7, value=row_data['date'])
        # Column H (Duplicate) is intentionally left empty

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 32
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14

    # Freeze the header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Total data rows: {len(all_rows)}')

create_initial()
