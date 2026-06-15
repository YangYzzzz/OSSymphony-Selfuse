"""
Reward Script: Sales Pipeline Duplicate Cleanup
Task ID: calc_sales_pipeline_duplicate_cleanup_049
Domain: libreoffice_calc

Task: Find and flag duplicate leads by matching on Company+Email using COUNTIFS.
Mark duplicates in column H, filter to show only 'Duplicate' rows, and remove
leads with blank Email column.

Scoring Rubric:
  Component 1: Blank email rows removed from the dataset (0.3 pts)
               Initial has 15 rows with blank Email; golden should have 0.
  Component 2: Column H filled with COUNTIFS-based IF formula for all data rows (0.4 pts)
               H2:H<last_row> must each contain =IF(COUNTIFS(...)>1,"Duplicate","Unique")
  Component 3: AutoFilter applied and filtered on column H to show only 'Duplicate' rows (0.3 pts)
               ws.auto_filter.ref must be set and filterColumn for col H must filter 'Duplicate'
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_pipeline_duplicate_cleanup_049'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the LeadList sheet exists
    if 'LeadList' not in wb.sheetnames:
        print("FAIL: 'LeadList' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['LeadList']

    # -----------------------------------------------------------------------
    # Component 1: Blank email rows removed (0.3 points)
    # Initial file had 15 rows with blank Email (column E).
    # After the task, all rows with blank Email must be deleted.
    # This FAILS on initial (has 15 blank emails) → PASSES on golden (0 blank emails).
    # -----------------------------------------------------------------------
    try:
        blank_email_count = 0
        total_data_rows = 0
        for row in range(2, ws.max_row + 1):
            email_val = ws.cell(row=row, column=5).value
            total_data_rows += 1
            if email_val is None or str(email_val).strip() == '':
                blank_email_count += 1

        if blank_email_count == 0 and total_data_rows > 0:
            print(f"PASS: Component 1 — No blank Email rows found (total data rows: {total_data_rows}) (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — Found {blank_email_count} rows with blank Email "
                  f"(expected 0). Task requires removing rows with blank Email.")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Column H filled with COUNTIFS-based IF formula (0.4 points)
    # All data rows (row 2 to max_row) must have a formula in column H that uses
    # COUNTIFS on Company (col D) and Email (col E) to identify duplicates.
    # The formula pattern: =IF(COUNTIFS($D$...:$D$...,D<row>,$E$...:$E$...,E<row>)>1,"Duplicate","Unique")
    # This FAILS on initial (H column is empty) → PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        formula_count = 0
        countifs_pattern_count = 0
        duplicate_unique_count = 0
        no_formula_count = 0
        last_data_row = ws.max_row

        for row in range(2, last_data_row + 1):
            h_val = ws.cell(row=row, column=8).value
            if h_val is None or not str(h_val).startswith('='):
                no_formula_count += 1
                continue
            formula_count += 1
            formula_upper = str(h_val).upper().replace(' ', '')
            # Check for COUNTIFS usage with columns D and E
            if 'COUNTIFS' in formula_upper and '$D$' in formula_upper and '$E$' in formula_upper:
                countifs_pattern_count += 1
            # Check for "Duplicate"/"Unique" label output
            if '"DUPLICATE"' in formula_upper and '"UNIQUE"' in formula_upper:
                duplicate_unique_count += 1

        total_rows = last_data_row - 1  # data rows count
        if total_rows > 0:
            formula_ratio = formula_count / total_rows
        else:
            formula_ratio = 0.0

        if (formula_count == total_rows and
                countifs_pattern_count == total_rows and
                duplicate_unique_count == total_rows):
            print(f"PASS: Component 2 — All {formula_count}/{total_rows} data rows in column H "
                  f"have COUNTIFS-based IF formula (0.4 pts)")
            total_score += 0.4
        elif formula_count > 0 and countifs_pattern_count > 0:
            # Partial: some rows have formulas but not all
            partial = round(0.4 * (countifs_pattern_count / total_rows), 2) if total_rows > 0 else 0.0
            print(f"PARTIAL: Component 2 — {countifs_pattern_count}/{total_rows} rows have correct "
                  f"COUNTIFS formula. Missing formulas in {no_formula_count} rows. "
                  f"Awarding {partial} pts")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Column H formulas not found or don't use COUNTIFS. "
                  f"formula_count={formula_count}, countifs_count={countifs_pattern_count}, "
                  f"total_rows={total_rows}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: AutoFilter applied and filtered on column H for 'Duplicate' (0.3 points)
    # The data range should have an AutoFilter set (auto_filter.ref is not None/empty),
    # and it must have a filter column definition on column H (colId=7, 0-indexed)
    # with filter value 'Duplicate'.
    # This FAILS on initial (no autofilter) → PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        has_autofilter = bool(ws.auto_filter.ref)
        has_h_filter = False
        filter_has_duplicate = False

        if has_autofilter:
            for fc in ws.auto_filter.filterColumn:
                # colId=7 means column H (0-indexed: A=0,B=1,...,H=7)
                if fc.colId == 7 and fc.filters:
                    has_h_filter = True
                    filter_values = [str(v).lower() for v in fc.filters.filter]
                    if 'duplicate' in filter_values:
                        filter_has_duplicate = True
                    break

        if has_autofilter and has_h_filter and filter_has_duplicate:
            print(f"PASS: Component 3 — AutoFilter set (ref={ws.auto_filter.ref}), "
                  f"column H filtered to 'Duplicate' (0.3 pts)")
            total_score += 0.3
        elif has_autofilter and has_h_filter:
            print(f"PARTIAL: Component 3 — AutoFilter set and H filter column exists, "
                  f"but filter value is not 'Duplicate'. Awarding 0.15 pts")
            total_score += 0.15
        elif has_autofilter:
            print(f"PARTIAL: Component 3 — AutoFilter is set (ref={ws.auto_filter.ref}), "
                  f"but no filter on column H found. Awarding 0.1 pts")
            total_score += 0.1
        else:
            print(f"FAIL: Component 3 — No AutoFilter found (auto_filter.ref={ws.auto_filter.ref}). "
                  f"Task requires AutoFilter with column H filtered to show 'Duplicate' rows.")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
