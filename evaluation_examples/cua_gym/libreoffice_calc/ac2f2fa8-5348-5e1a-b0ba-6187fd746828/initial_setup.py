"""
Initial Setup: Set conditional formatting for bottom 10% scores with orange background
Task ID: calc_gg3_030
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_030'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Scores ---
    ws = wb.active
    ws.title = 'Scores'

    # Headers
    headers = ['Student ID', 'First Name', 'Last Name', 'Department', 'Assessment Date', 'Score']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 10

    # Student data - 40 rows with realistic names and varied scores
    students = [
        ['S1001', 'Sarah', 'Chen', 'Computer Science', '2025-09-15', 87],
        ['S1002', 'Marcus', 'Johnson', 'Mathematics', '2025-09-15', 74],
        ['S1003', 'Emily', 'Rodriguez', 'Physics', '2025-09-16', 91],
        ['S1004', 'James', 'O\'Brien', 'Computer Science', '2025-09-16', 68],
        ['S1005', 'Aisha', 'Patel', 'Chemistry', '2025-09-17', 95],
        ['S1006', 'David', 'Kim', 'Mathematics', '2025-09-17', 82],
        ['S1007', 'Sofia', 'Andersson', 'Biology', '2025-09-18', 45],
        ['S1008', 'Wei', 'Zhang', 'Computer Science', '2025-09-18', 93],
        ['S1009', 'Olivia', 'Thompson', 'Physics', '2025-09-19', 76],
        ['S1010', 'Liam', 'Murphy', 'Chemistry', '2025-09-19', 88],
        ['S1011', 'Fatima', 'Hassan', 'Mathematics', '2025-09-20', 71],
        ['S1012', 'Noah', 'Garcia', 'Biology', '2025-09-20', 84],
        ['S1013', 'Yuki', 'Tanaka', 'Computer Science', '2025-09-21', 92],
        ['S1014', 'Isabella', 'Rossi', 'Physics', '2025-09-21', 63],
        ['S1015', 'Ethan', 'Williams', 'Chemistry', '2025-09-22', 79],
        ['S1016', 'Priya', 'Sharma', 'Mathematics', '2025-09-22', 86],
        ['S1017', 'Lucas', 'Dubois', 'Biology', '2025-09-23', 41],
        ['S1018', 'Amara', 'Okafor', 'Computer Science', '2025-09-23', 90],
        ['S1019', 'Benjamin', 'Clarke', 'Physics', '2025-09-24', 77],
        ['S1020', 'Mei', 'Liu', 'Chemistry', '2025-09-24', 94],
        ['S1021', 'Alexander', 'Petrov', 'Mathematics', '2025-09-25', 69],
        ['S1022', 'Chloe', 'Martin', 'Biology', '2025-09-25', 85],
        ['S1023', 'Daniel', 'Nakamura', 'Computer Science', '2025-09-26', 73],
        ['S1024', 'Zara', 'Ahmed', 'Physics', '2025-09-26', 96],
        ['S1025', 'Ryan', 'Kowalski', 'Chemistry', '2025-09-27', 81],
        ['S1026', 'Hannah', 'Svensson', 'Mathematics', '2025-09-27', 88],
        ['S1027', 'Omar', 'Diaz', 'Biology', '2025-09-28', 52],
        ['S1028', 'Grace', 'Taylor', 'Computer Science', '2025-09-28', 90],
        ['S1029', 'Adrian', 'Muller', 'Physics', '2025-09-29', 78],
        ['S1030', 'Leila', 'Fernandez', 'Chemistry', '2025-09-29', 83],
        ['S1031', 'Thomas', 'Nguyen', 'Mathematics', '2025-09-30', 67],
        ['S1032', 'Natalie', 'Brown', 'Biology', '2025-09-30', 91],
        ['S1033', 'Kevin', 'Schmidt', 'Computer Science', '2025-10-01', 75],
        ['S1034', 'Rina', 'Watanabe', 'Physics', '2025-10-01', 89],
        ['S1035', 'Samuel', 'Costa', 'Chemistry', '2025-10-02', 43],
        ['S1036', 'Victoria', 'Lee', 'Mathematics', '2025-10-02', 87],
        ['S1037', 'Hassan', 'Ali', 'Biology', '2025-10-03', 72],
        ['S1038', 'Emma', 'Johansson', 'Computer Science', '2025-10-03', 93],
        ['S1039', 'Carlos', 'Reyes', 'Physics', '2025-10-04', 58],
        ['S1040', 'Lily', 'Wang', 'Chemistry', '2025-10-04', 80],
    ]

    for r, row_data in enumerate(students, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # NO conditional formatting in initial state - that's the task

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
