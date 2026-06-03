"""
Reward Script: Verify Bottom 10% conditional formatting on Scores sheet
Task ID: calc_gg3_030
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): CF rule exists targeting range F2:F41
  Component 2 (0.35): Rule is top10 type with Bottom=True and Percent=True
  Component 3 (0.15): Rank/threshold value is 10
  Component 4 (0.20): Fill color is orange (FFFFA500)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_030'


def persist_app_state(domain: str):
    """Save any unsaved LibreOffice edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify that a Bottom 10% conditional formatting rule with orange fill
    has been applied to F2:F41 on the 'Scores' sheet.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check 'Scores' sheet exists (precondition gate)
    if 'Scores' not in wb.sheetnames:
        print(f"CRITICAL: 'Scores' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Scores']

    # Gather all conditional formatting rules
    cf_rules_list = []
    for cf in ws.conditional_formatting:
        cf_range = str(cf)
        for rule in cf.rules:
            cf_rules_list.append((cf_range, rule))

    print(f"INFO: Found {len(cf_rules_list)} conditional formatting rule(s)")

    # Find the matching rule on F2:F41 (or containing that range)
    target_rule = None
    target_range = None
    for cf_range_str, rule in cf_rules_list:
        # Check if the CF range covers F2:F41
        # The range string from openpyxl looks like "<ConditionalFormatting F2:F41>"
        # or just the cell range portion. Normalize.
        range_clean = cf_range_str.replace('<ConditionalFormatting ', '').replace('>', '').strip()
        # Accept exact match or ranges that contain F2:F41
        if 'F2:F41' in range_clean or 'F2:F41' in cf_range_str:
            target_rule = rule
            target_range = range_clean
            break

    # Component 1: CF rule exists on range F2:F41 (0.30 points)
    try:
        if target_rule is not None:
            print(f"PASS: Component 1 — CF rule found on range {target_range} (0.30 pts)")
            total_score += 0.30
        else:
            # Maybe a rule exists but on a slightly different range representation
            # Check for any top10 rule on the sheet that relates to column F
            for cf_range_str, rule in cf_rules_list:
                range_clean = cf_range_str.replace('<ConditionalFormatting ', '').replace('>', '').strip()
                if rule.type == 'top10' and 'F' in range_clean:
                    target_rule = rule
                    target_range = range_clean
                    print(f"PASS: Component 1 — CF top10 rule found on range {target_range} (close match) (0.30 pts)")
                    total_score += 0.30
                    break
            if target_rule is None:
                print(f"FAIL: Component 1 — No CF rule found targeting F2:F41. Total rules: {len(cf_rules_list)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # If no rule found at all, remaining checks cannot pass
    if target_rule is None:
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Component 2: Rule type is top10 with Bottom=True and Percent=True (0.35 points)
    try:
        is_top10 = (target_rule.type == 'top10')
        is_bottom = getattr(target_rule, 'bottom', False)
        is_percent = getattr(target_rule, 'percent', False)

        if is_top10 and is_bottom and is_percent:
            print(f"PASS: Component 2 — Rule type=top10, bottom=True, percent=True (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 2 — type={target_rule.type}, bottom={is_bottom}, percent={is_percent}")
            # Partial: if type is correct but missing bottom/percent flags
            if is_top10:
                partial = 0.15
                print(f"  Partial: type=top10 correct (+{partial} pts)")
                total_score += partial
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Rank is 10 (the 10% threshold) (0.15 points)
    try:
        rank_val = getattr(target_rule, 'rank', None)
        if rank_val is not None and int(rank_val) == 10:
            print(f"PASS: Component 3 — Rank/threshold = {rank_val} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 — Expected rank=10, found: {rank_val}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Fill color is orange (0.20 points)
    try:
        dxf = target_rule.dxf
        if dxf and dxf.fill:
            fg_rgb = None
            bg_rgb = None
            if dxf.fill.fgColor and dxf.fill.fgColor.rgb:
                fg_rgb = str(dxf.fill.fgColor.rgb)
            if dxf.fill.bgColor and dxf.fill.bgColor.rgb:
                bg_rgb = str(dxf.fill.bgColor.rgb)

            # Orange color: FFA500 in various ARGB forms
            # Accept common orange variants
            orange_variants = [
                'FFFFA500',  # exact orange
                'FFFF8C00',  # dark orange
                'FFFF6600',  # deeper orange
                'FFFFC000',  # gold-orange
                'FFED7D31',  # LibreOffice default orange
                'FFFF9900',  # web orange
            ]

            # Check if either fg or bg color is orange-ish
            color_found = fg_rgb or bg_rgb
            is_orange = False
            if color_found:
                # Check exact match first
                if color_found in orange_variants:
                    is_orange = True
                else:
                    # Check if the color is in the orange hue range (R high, G medium, B low)
                    try:
                        hex_color = color_found[-6:]  # last 6 chars (RGB without alpha)
                        r = int(hex_color[0:2], 16)
                        g = int(hex_color[2:4], 16)
                        b = int(hex_color[4:6], 16)
                        # Orange: R >= 200, 80 <= G <= 200, B <= 100
                        if r >= 200 and 60 <= g <= 210 and b <= 100:
                            is_orange = True
                    except Exception:
                        pass

            if is_orange:
                print(f"PASS: Component 4 — Orange fill color: {color_found} (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 4 — Expected orange fill, found fg={fg_rgb}, bg={bg_rgb}")
        else:
            print(f"FAIL: Component 4 — No fill defined in the conditional formatting rule")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
