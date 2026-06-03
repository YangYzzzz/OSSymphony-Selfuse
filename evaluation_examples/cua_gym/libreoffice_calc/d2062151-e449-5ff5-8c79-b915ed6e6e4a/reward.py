"""
Reward Script: Protected sales quote template with locked formula cells,
               unlocked input cells, and sheet protection with password.
Task ID: calc_sales_086
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Line-item formulas E7:E11 = Cx*Dx, locked
  Component 2 (0.25): Summary formulas E13, E14, E15, locked
  Component 3 (0.25): Input cells (B3, B4, A7:D11) are unlocked
  Component 4 (0.25): Sheet protection enabled with password
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_086'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists
    if 'Quote' not in wb.sheetnames:
        print("FAIL: 'Quote' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Quote']

    # Component 1: Line-item formulas E7:E11 = Cx*Dx and locked (0.25 points)
    # Initial env has NO formulas in these cells, so this only passes on golden.
    try:
        formula_cells = {
            'E7': '=C7*D7',
            'E8': '=C8*D8',
            'E9': '=C9*D9',
            'E10': '=C10*D10',
            'E11': '=C11*D11',
        }
        formulas_ok = 0
        for coord, expected in formula_cells.items():
            cell = ws[coord]
            val = cell.value
            if val is not None and isinstance(val, str):
                # Normalize: strip spaces, uppercase
                if val.upper().replace(" ", "") == expected.upper().replace(" ", ""):
                    if cell.protection.locked:
                        formulas_ok += 1
                    else:
                        print(f"FAIL: {coord} has correct formula but is NOT locked")
                else:
                    print(f"FAIL: {coord} expected {expected}, found {val}")
            else:
                print(f"FAIL: {coord} expected formula {expected}, found {repr(val)}")

        if formulas_ok == 5:
            print(f"PASS: Component 1 -- All 5 line-item formulas present and locked (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 -- {formulas_ok}/5 line-item formulas correct")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Summary formulas E13, E14, E15 and locked (0.25 points)
    # Initial env has NO formulas in these cells.
    try:
        summary_formulas = {
            'E13': '=SUM(E7:E11)',
            'E14': '=E13*0.08',
            'E15': '=E13+E14',
        }
        summary_ok = 0
        for coord, expected in summary_formulas.items():
            cell = ws[coord]
            val = cell.value
            if val is not None and isinstance(val, str):
                if val.upper().replace(" ", "") == expected.upper().replace(" ", ""):
                    if cell.protection.locked:
                        summary_ok += 1
                    else:
                        print(f"FAIL: {coord} has correct formula but is NOT locked")
                else:
                    print(f"FAIL: {coord} expected {expected}, found {val}")
            else:
                print(f"FAIL: {coord} expected formula {expected}, found {repr(val)}")

        if summary_ok == 3:
            print(f"PASS: Component 2 -- All 3 summary formulas present and locked (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 -- {summary_ok}/3 summary formulas correct")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Input cells are unlocked (0.25 points)
    # Initial env has all cells LOCKED (default), so this only passes on golden.
    try:
        input_cells = ['B3', 'B4']
        # A7:D11 range
        for row in range(7, 12):
            for col_letter in ['A', 'B', 'C', 'D']:
                input_cells.append(f'{col_letter}{row}')

        unlocked_count = 0
        for coord in input_cells:
            cell = ws[coord]
            if not cell.protection.locked:
                unlocked_count += 1
            else:
                print(f"FAIL: Input cell {coord} is locked (should be unlocked)")

        total_input = len(input_cells)  # 22 cells
        if unlocked_count == total_input:
            print(f"PASS: Component 3 -- All {total_input} input cells are unlocked (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 3 -- {unlocked_count}/{total_input} input cells unlocked")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Sheet protection enabled with password (0.25 points)
    # Initial env has protection DISABLED.
    try:
        if ws.protection.sheet:
            # Check that a password hash is set (non-empty)
            if ws.protection.password:
                print(f"PASS: Component 4 -- Sheet protection enabled with password (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 4 -- Sheet protection enabled but NO password set")
        else:
            print(f"FAIL: Component 4 -- Sheet protection is NOT enabled")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
