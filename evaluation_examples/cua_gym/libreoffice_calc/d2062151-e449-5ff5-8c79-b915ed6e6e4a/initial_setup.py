"""
Initial Setup: Create a sales quote template with headers and structure, no formulas or protection.
Task ID: calc_sales_086
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_086'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quote"

    # --- Title row: merged A1:E1 ---
    ws.merge_cells("A1:E1")
    ws["A1"] = "SALES QUOTE"
    ws["A1"].font = Font(name="Arial", size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # --- Customer and Date labels ---
    ws["A3"] = "Customer:"
    ws["A3"].font = Font(name="Arial", size=11, bold=True)
    ws["A4"] = "Date:"
    ws["A4"].font = Font(name="Arial", size=11, bold=True)

    # B3 and B4 are left empty (input cells for agent)

    # --- Column headers row 6 ---
    headers = ["Item", "Description", "Qty", "Unit Price", "Total"]
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Data rows 7-11: empty but with borders for structure ---
    for row in range(7, 12):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

    # --- Summary labels ---
    ws["A13"] = "Subtotal"
    ws["A13"].font = Font(name="Arial", size=11, bold=True)
    ws["A14"] = "Tax (8%)"
    ws["A14"].font = Font(name="Arial", size=11, bold=True)
    ws["A15"] = "Total"
    ws["A15"].font = Font(name="Arial", size=12, bold=True)

    # Summary cells E13:E15 are empty (no formulas yet)
    for row in range(13, 16):
        ws.cell(row=row, column=5).border = thin_border

    # --- Column widths ---
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    # --- Number format hints for Qty, Unit Price, Total columns ---
    for row in range(7, 12):
        ws.cell(row=row, column=3).number_format = '0'          # Qty as integer
        ws.cell(row=row, column=4).number_format = '$#,##0.00'  # Unit Price
        ws.cell(row=row, column=5).number_format = '$#,##0.00'  # Total

    for row in range(13, 16):
        ws.cell(row=row, column=5).number_format = '$#,##0.00'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
