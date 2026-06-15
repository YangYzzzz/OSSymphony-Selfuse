"""
Reward Script: Benefits enrollment tracker with dropdowns and VLOOKUP
Task ID: calc_hr_benefit_enrollment_017
Domain: libreoffice_calc
Scoring:
  - Component 1: Health Plan dropdown validation on C2:C112 (0.25 pts)
  - Component 2: Dental dropdown validation on D2:D112 (0.15 pts)
  - Component 3: Vision dropdown validation on E2:E112 (0.15 pts)
  - Component 4: VLOOKUP formula in F2:F112 (0.35 pts)
  - Component 5: Currency number format on F2:F112 (0.10 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_benefit_enrollment_017'


def normalize_formula(formula):
    """Normalize formula for comparison: uppercase, remove spaces."""
    if not isinstance(formula, str):
        return ''
    return formula.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Ensure Enrollment sheet exists
    if 'Enrollment' not in wb.sheetnames:
        print("FAIL: Sheet 'Enrollment' not found in workbook")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws = wb['Enrollment']
    dvs = ws.data_validations.dataValidation

    # --- Component 1: Health Plan dropdown validation on C2:C112 (0.25 pts) ---
    # Initial file has NO validation; golden adds 'Basic,Standard,Premium' list on C2:C112
    try:
        health_dv_found = False
        for dv in dvs:
            if dv.type == 'list':
                formula = dv.formula1
                if formula and 'Basic' in formula and 'Standard' in formula and 'Premium' in formula:
                    # Check that the sqref covers C2:C112
                    sqref_str = str(dv.sqref)
                    if 'C2:C112' in sqref_str or sqref_str == 'C2:C112':
                        health_dv_found = True
                        print(f"PASS: Component 1 — Health Plan dropdown 'Basic,Standard,Premium' on C2:C112 (0.25 pts)")
                        total_score += 0.25
                        break
        if not health_dv_found:
            # Check if any health plan validation exists with slightly different range
            for dv in dvs:
                if dv.type == 'list' and dv.formula1 and 'Basic' in dv.formula1:
                    sqref_str = str(dv.sqref)
                    print(f"FAIL: Component 1 — Health Plan validation found but range is {sqref_str}, expected C2:C112. formula1={dv.formula1}")
                    break
            else:
                print(f"FAIL: Component 1 — No Health Plan dropdown (Basic,Standard,Premium) validation on C2:C112 found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # --- Component 2: Dental dropdown validation on D2:D112 (0.15 pts) ---
    # Initial file has NO validation; golden adds 'Yes,No' list on D2:D112
    try:
        dental_dv_found = False
        for dv in dvs:
            if dv.type == 'list':
                formula = dv.formula1
                sqref_str = str(dv.sqref)
                if formula and 'Yes' in formula and 'No' in formula:
                    if 'D2:D112' in sqref_str or sqref_str == 'D2:D112':
                        dental_dv_found = True
                        print(f"PASS: Component 2 — Dental dropdown 'Yes,No' on D2:D112 (0.15 pts)")
                        total_score += 0.15
                        break
        if not dental_dv_found:
            for dv in dvs:
                if dv.type == 'list' and dv.formula1 and 'Yes' in dv.formula1:
                    sqref_str = str(dv.sqref)
                    if 'D' in sqref_str:
                        print(f"FAIL: Component 2 — Dental validation found but range is {sqref_str}, expected D2:D112")
                        break
            else:
                print(f"FAIL: Component 2 — No Dental dropdown (Yes,No) validation on D2:D112 found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # --- Component 3: Vision dropdown validation on E2:E112 (0.15 pts) ---
    # Initial file has NO validation; golden adds 'Yes,No' list on E2:E112
    try:
        vision_dv_found = False
        for dv in dvs:
            if dv.type == 'list':
                formula = dv.formula1
                sqref_str = str(dv.sqref)
                if formula and 'Yes' in formula and 'No' in formula:
                    if 'E2:E112' in sqref_str or sqref_str == 'E2:E112':
                        vision_dv_found = True
                        print(f"PASS: Component 3 — Vision dropdown 'Yes,No' on E2:E112 (0.15 pts)")
                        total_score += 0.15
                        break
        if not vision_dv_found:
            for dv in dvs:
                if dv.type == 'list' and dv.formula1 and 'Yes' in dv.formula1:
                    sqref_str = str(dv.sqref)
                    if 'E' in sqref_str:
                        print(f"FAIL: Component 3 — Vision validation found but range is {sqref_str}, expected E2:E112")
                        break
            else:
                print(f"FAIL: Component 3 — No Vision dropdown (Yes,No) validation on E2:E112 found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # --- Component 4: VLOOKUP formula in F2:F112 (0.35 pts) ---
    # Initial file has None in F2:F112; golden has VLOOKUP formula in all rows
    # Check key formula elements: VLOOKUP, Benefits Costs, IF, Dental B5, Vision B6
    try:
        formula_rows_correct = 0
        formula_rows_checked = 0
        sample_rows = list(range(2, 12)) + [50, 112]  # sample first 10 rows + mid + last

        for row in sample_rows:
            cell_val = ws.cell(row=row, column=6).value
            if cell_val is not None:
                formula_rows_checked += 1
                norm = normalize_formula(str(cell_val))
                # Check for VLOOKUP referencing C column and Benefits Costs
                has_vlookup = 'VLOOKUP(' in norm
                has_benefits_ref = "'BENEFITSCOSTS'" in norm or 'BENEFITSCOSTS' in norm
                has_if_dental = 'D' + str(row) in norm and 'IF(' in norm
                has_if_vision = 'E' + str(row) in norm
                has_b5_ref = '$B$5' in norm
                has_b6_ref = '$B$6' in norm

                if has_vlookup and has_benefits_ref and has_if_dental and has_if_vision:
                    formula_rows_correct += 1

        # Need most sampled rows to have correct formula
        if formula_rows_checked == 0:
            print(f"FAIL: Component 4 — F2:F112 cells are empty (no formulas found)")
        elif formula_rows_correct >= len(sample_rows) * 0.8:
            print(f"PASS: Component 4 — VLOOKUP formula found in F2:F112 ({formula_rows_correct}/{formula_rows_checked} sampled rows correct) (0.35 pts)")
            total_score += 0.35
        else:
            # Partial: formula exists but may not cover all rows
            # Check if at least the first row has valid formula
            f2_val = ws.cell(row=2, column=6).value
            if f2_val and 'VLOOKUP' in str(f2_val).upper():
                print(f"FAIL: Component 4 — VLOOKUP formula present in F2 but only {formula_rows_correct}/{formula_rows_checked} sampled rows correct")
            else:
                print(f"FAIL: Component 4 — F2 formula is {repr(f2_val)}, expected VLOOKUP formula")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # --- Component 5: Currency number format on F2:F112 (0.10 pts) ---
    # Initial file has 'General' format; golden has '$#,##0.00' format
    try:
        currency_count = 0
        sample_f_rows = list(range(2, 7)) + [50, 112]
        for row in sample_f_rows:
            fmt = ws.cell(row=row, column=6).number_format
            if fmt and ('$' in fmt or '#,##0.00' in fmt):
                currency_count += 1

        if currency_count >= len(sample_f_rows) * 0.8:
            print(f"PASS: Component 5 — F2:F112 formatted as currency ({currency_count}/{len(sample_f_rows)} sampled rows have currency format) (0.10 pts)")
            total_score += 0.10
        else:
            # Check F2 specifically for diagnostic info
            f2_fmt = ws.cell(row=2, column=6).number_format
            print(f"FAIL: Component 5 — Expected currency format $#,##0.00, only {currency_count}/{len(sample_f_rows)} sampled rows correct. F2 format={repr(f2_fmt)}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
