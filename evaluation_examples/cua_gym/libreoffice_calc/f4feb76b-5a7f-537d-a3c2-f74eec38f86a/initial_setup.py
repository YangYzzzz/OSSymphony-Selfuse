"""
Initial Setup: Benefits Enrollment Tracker
Task ID: calc_hr_benefit_enrollment_017
Domain: libreoffice_calc

Creates a spreadsheet with:
- 'Enrollment' sheet: employee data rows 2-112, NO validation in C/D/E, NO formulas in F
- 'Benefits Costs' sheet: plan cost reference table
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_benefit_enrollment_017'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Enrollment ---
    ws_enroll = wb.active
    ws_enroll.title = 'Enrollment'

    # Headers
    headers = ['Emp ID', 'Name', 'Health Plan', 'Dental', 'Vision', 'Monthly Deduction']
    for col, h in enumerate(headers, 1):
        ws_enroll.cell(row=1, column=col, value=h)

    # Employee data - 111 rows (rows 2-112), realistic names and IDs
    employees = [
        ('E001', 'Sarah Chen'),
        ('E002', 'Marcus Johnson'),
        ('E003', 'Priya Patel'),
        ('E004', 'James O\'Brien'),
        ('E005', 'Aisha Williams'),
        ('E006', 'Derek Hoffman'),
        ('E007', 'Mei-Ling Torres'),
        ('E008', 'Robert Castellano'),
        ('E009', 'Fatima Al-Rashid'),
        ('E010', 'Thomas Nguyen'),
        ('E011', 'Gabrielle Fontaine'),
        ('E012', 'Kevin Park'),
        ('E013', 'Natasha Ivanova'),
        ('E014', 'Carlos Rivera'),
        ('E015', 'Jennifer Blackwood'),
        ('E016', 'Samuel Okafor'),
        ('E017', 'Lisa Montgomery'),
        ('E018', 'David Kowalski'),
        ('E019', 'Yuki Tanaka'),
        ('E020', 'Patricia Summers'),
        ('E021', 'Ravi Krishnamurthy'),
        ('E022', 'Amanda Fletcher'),
        ('E023', 'Winston Hayes'),
        ('E024', 'Ingrid Sorensen'),
        ('E025', 'Mohammed Al-Hassan'),
        ('E026', 'Rebecca Thornton'),
        ('E027', 'Philip Mercer'),
        ('E028', 'Diana Christodoulou'),
        ('E029', 'Anthony Reyes'),
        ('E030', 'Claire Beaumont'),
        ('E031', 'Nathan Goldberg'),
        ('E032', 'Sonia Kapoor'),
        ('E033', 'Gregory Marsh'),
        ('E034', 'Leila Ahmadi'),
        ('E035', 'Brandon Whitfield'),
        ('E036', 'Chantal Dubois'),
        ('E037', 'Isaac Zimmerman'),
        ('E038', 'Rosa Gutierrez'),
        ('E039', 'Peter Larsson'),
        ('E040', 'Vanessa Osei'),
        ('E041', 'Mitchell Crawford'),
        ('E042', 'Amelia Johansson'),
        ('E043', 'Felix Schneider'),
        ('E044', 'Tamara Petrov'),
        ('E045', 'Darius Washington'),
        ('E046', 'Simone Laurent'),
        ('E047', 'Walter Eriksson'),
        ('E048', 'Beatrice Nkosi'),
        ('E049', 'Simon Harrington'),
        ('E050', 'Nadia Kowalczyk'),
        ('E051', 'Jerome Baptiste'),
        ('E052', 'Vera Stankovic'),
        ('E053', 'Leonard Ferreira'),
        ('E054', 'Helena Makinen'),
        ('E055', 'Oscar Mendoza'),
        ('E056', 'Anastasia Volkov'),
        ('E057', 'Brendan Fitzgerald'),
        ('E058', 'Yasmin Chowdhury'),
        ('E059', 'Stefan Muller'),
        ('E060', 'Celeste Diallo'),
        ('E061', 'Howard Bergmann'),
        ('E062', 'Xiomara Vega'),
        ('E063', 'Alistair MacPherson'),
        ('E064', 'Keiko Watanabe'),
        ('E065', 'Eduardo Almeida'),
        ('E066', 'Monique Rousseau'),
        ('E067', 'Liam O\'Sullivan'),
        ('E068', 'Miriam Svensson'),
        ('E069', 'Tobias Braun'),
        ('E070', 'Esperanza Lozano'),
        ('E071', 'Cornelius van der Berg'),
        ('E072', 'Adaeze Okonkwo'),
        ('E073', 'Rudolf Holzer'),
        ('E074', 'Jasmine Archibald'),
        ('E075', 'Piotr Wisniewski'),
        ('E076', 'Ingrid Halvorsen'),
        ('E077', 'Terrence Boateng'),
        ('E078', 'Sofia Castellanos'),
        ('E079', 'Eugene Nakamura'),
        ('E080', 'Harriet Blackburn'),
        ('E081', 'Kwame Asante'),
        ('E082', 'Olga Romanova'),
        ('E083', 'Duncan McAllister'),
        ('E084', 'Laleh Shirazi'),
        ('E085', 'Vicente Herrera'),
        ('E086', 'Brigitte Morel'),
        ('E087', 'Frederic Dupont'),
        ('E088', 'Amara Diarra'),
        ('E089', 'Aleksei Petrov'),
        ('E090', 'Constance Whitmore'),
        ('E091', 'Ibrahim Oduya'),
        ('E092', 'Soledad Ramos'),
        ('E093', 'Henrik Lindqvist'),
        ('E094', 'Annette Brauer'),
        ('E095', 'Desmond Abara'),
        ('E096', 'Valentina Cruz'),
        ('E097', 'Matteo Ricci'),
        ('E098', 'Winnie Adjei'),
        ('E099', 'Javier Morales'),
        ('E100', 'Cecile Girard'),
        ('E101', 'Obinna Eze'),
        ('E102', 'Ekaterina Sokolova'),
        ('E103', 'Malcolm Shaw'),
        ('E104', 'Farida Yusupova'),
        ('E105', 'Thierry Lemaire'),
        ('E106', 'Adwoa Mensah'),
        ('E107', 'Dmitri Volkov'),
        ('E108', 'Penelope Hart'),
        ('E109', 'Rajesh Subramaniam'),
        ('E110', 'Nathalie Boivin'),
        ('E111', 'Emeka Obi'),
    ]

    # Health plan options to distribute across employees
    health_plans = ['Basic', 'Standard', 'Premium']
    dental_opts = ['Yes', 'No']
    vision_opts = ['Yes', 'No']

    for i, (emp_id, name) in enumerate(employees):
        row = i + 2
        # Distribute health plans realistically
        plan = health_plans[i % 3]
        dental = dental_opts[i % 2]
        vision = vision_opts[(i // 2) % 2]

        ws_enroll.cell(row=row, column=1, value=emp_id)
        ws_enroll.cell(row=row, column=2, value=name)
        ws_enroll.cell(row=row, column=3, value=plan)
        ws_enroll.cell(row=row, column=4, value=dental)
        ws_enroll.cell(row=row, column=5, value=vision)
        # Column F (Monthly Deduction): leave empty — no formula yet
        # (no value set for column 6)

    # --- Sheet 2: Benefits Costs ---
    ws_costs = wb.create_sheet('Benefits Costs')

    # Headers
    ws_costs.cell(row=1, column=1, value='Plan')
    ws_costs.cell(row=1, column=2, value='Monthly Cost')

    # Cost data
    cost_data = [
        ('Basic', 150),
        ('Standard', 275),
        ('Premium', 420),
        ('Dental', 25),
        ('Vision', 15),
    ]
    for r, (plan, cost) in enumerate(cost_data, 2):
        ws_costs.cell(row=r, column=1, value=plan)
        ws_costs.cell(row=r, column=2, value=cost)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Enrollment sheet: {ws_enroll.max_row} rows, {ws_enroll.max_column} columns')
    print(f'  Benefits Costs sheet: {ws_costs.max_row} rows')
    print('  No validation in C/D/E, no formulas in F (as required)')


create_initial()
