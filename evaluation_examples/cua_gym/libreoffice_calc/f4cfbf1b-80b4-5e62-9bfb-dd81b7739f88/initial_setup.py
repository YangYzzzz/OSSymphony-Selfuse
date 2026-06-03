"""
Initial Setup: Create spreadsheet with student exam score distribution data
Task ID: calc_chart_histogram_072
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_histogram_072'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: ScoreDistribution ---
    ws = wb.active
    ws.title = 'ScoreDistribution'

    # Headers
    ws['A1'] = 'Score Range'
    ws['B1'] = 'Count'

    # Data rows — exact values from task context
    data = [
        ('0-10',   2),
        ('11-20',  4),
        ('21-30',  8),
        ('31-40',  15),
        ('41-50',  28),
        ('51-60',  42),
        ('61-70',  58),
        ('71-80',  47),
        ('81-90',  31),
        ('91-100', 14),
    ]

    for r, (score_range, count) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=score_range)
        ws.cell(row=r, column=2, value=count)

    # No charts in initial file (task requires agent to create the chart)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
