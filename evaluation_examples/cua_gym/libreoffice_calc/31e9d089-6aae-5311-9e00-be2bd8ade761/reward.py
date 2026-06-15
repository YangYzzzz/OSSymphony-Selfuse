"""
Reward Script: Pivot table showing top 5 products by revenue, sorted descending
Task ID: calc_pivot_054
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Pivot sheet exists
  Component 2 (0.20): Correct headers (Product + Revenue-related)
  Component 3 (0.25): Exactly 5 product rows sorted descending by revenue
  Component 4 (0.30): Correct top-5 product names and revenue values
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_054'

# Expected top-5 products and revenues from task context
EXPECTED_TOP5 = [
    ('Prod_C', 42000),
    ('Prod_H', 38000),
    ('Prod_A', 35000),
    ('Prod_K', 32000),
    ('Prod_E', 29000),
]


def persist_app_state(domain: str):
    """Best-effort save of any unsaved GUI state."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Pivot sheet exists (0.25 points)
    # Initial file has only 'ProductRevenue'; golden adds a pivot sheet.
    # We look for any sheet OTHER than 'ProductRevenue' that could be the pivot.
    try:
        non_source_sheets = [s for s in wb.sheetnames if s != 'ProductRevenue']
        if len(non_source_sheets) >= 1:
            pivot_sheet_name = non_source_sheets[0]
            print(f"PASS: Component 1 -- Found pivot sheet '{pivot_sheet_name}' (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 -- No additional sheet found besides 'ProductRevenue'. Sheets: {wb.sheetnames}")
            final_score = min(total_score, 1.0)
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {final_score}")
            return final_score
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    ws = wb[pivot_sheet_name]

    # Component 2: Headers contain Product and Revenue-related column (0.20 points)
    try:
        header_a = ws.cell(row=1, column=1).value
        header_b = ws.cell(row=1, column=2).value
        header_a_str = str(header_a).lower().strip() if header_a else ''
        header_b_str = str(header_b).lower().strip() if header_b else ''

        has_product_header = 'product' in header_a_str
        has_revenue_header = 'revenue' in header_b_str

        if has_product_header and has_revenue_header:
            print(f"PASS: Component 2 -- Headers: '{header_a}', '{header_b}' (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 -- Expected Product/Revenue headers, found: '{header_a}', '{header_b}'")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Read all data rows (skip header, skip total row if present)
    data_rows = []
    try:
        for r in range(2, ws.max_row + 1):
            product = ws.cell(row=r, column=1).value
            revenue = ws.cell(row=r, column=2).value
            if product is not None and str(product).strip().lower() != 'total':
                try:
                    rev_val = float(revenue) if revenue is not None else None
                except (ValueError, TypeError):
                    rev_val = None
                data_rows.append((str(product).strip(), rev_val))
        print(f"INFO: Found {len(data_rows)} product rows: {data_rows}")
    except Exception as e:
        print(f"ERROR: Reading data rows -- {e}")

    # Component 3: Exactly 5 product rows sorted descending by revenue (0.25 points)
    try:
        count_ok = len(data_rows) == 5
        # Check descending sort: all consecutive pairs must be non-increasing
        sorted_desc = (
            len(data_rows) >= 2
            and all(
                data_rows[i][1] is not None
                and data_rows[i + 1][1] is not None
                and data_rows[i][1] >= data_rows[i + 1][1]
                for i in range(len(data_rows) - 1)
            )
        ) if len(data_rows) >= 2 else (len(data_rows) == 1)

        if count_ok and sorted_desc:
            print(f"PASS: Component 3 -- 5 rows, sorted descending (0.25 pts)")
            total_score += 0.25
        elif count_ok and not sorted_desc:
            print(f"PARTIAL: Component 3 -- 5 rows but NOT sorted descending (0.10 pts)")
            total_score += 0.10
        elif not count_ok and sorted_desc and len(data_rows) > 0:
            print(f"PARTIAL: Component 3 -- {len(data_rows)} rows (expected 5), but sorted (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 -- {len(data_rows)} rows, sorted_desc={sorted_desc}")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Correct top-5 product names and revenue values (0.30 points)
    # Check each expected product-revenue pair
    try:
        matches = 0
        for expected_product, expected_revenue in EXPECTED_TOP5:
            for actual_product, actual_revenue in data_rows:
                if actual_product == expected_product:
                    if actual_revenue is not None and abs(actual_revenue - expected_revenue) <= 1.0:
                        matches += 1
                    else:
                        print(f"  MISMATCH: {expected_product} expected revenue={expected_revenue}, found={actual_revenue}")
                    break
            else:
                print(f"  MISSING: {expected_product} not found in pivot data")

        if matches == 5:
            print(f"PASS: Component 4 -- All 5 products match with correct revenue (0.30 pts)")
            total_score += 0.30
        elif matches > 0:
            partial = round(0.30 * (matches / 5), 2)
            print(f"PARTIAL: Component 4 -- {matches}/5 products match ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 -- No products match expected values")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
