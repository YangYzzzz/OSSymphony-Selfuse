"""
Initial Setup: Create ProductRevenue spreadsheet with 400 transactions
Task ID: calc_pivot_054
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time
import openpyxl
from datetime import datetime, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_054'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(54)

    # 15 products: Prod_A through Prod_O
    products = [f'Prod_{chr(65+i)}' for i in range(15)]
    regions = ['North', 'South', 'East', 'West']

    # Target revenue sums for top 5:
    # Prod_C=42000, Prod_H=38000, Prod_A=35000, Prod_K=32000, Prod_E=29000
    # Remaining 10 products get lower totals (spread across ~10000-25000 each)

    # We'll assign each product a target total and a transaction count,
    # then distribute revenue across those transactions.
    # Total 400 transactions across 15 products.

    product_targets = {
        'Prod_A': 35000,
        'Prod_B': 18500,
        'Prod_C': 42000,
        'Prod_D': 15200,
        'Prod_E': 29000,
        'Prod_F': 21300,
        'Prod_G': 12800,
        'Prod_H': 38000,
        'Prod_I': 16700,
        'Prod_J': 10500,
        'Prod_K': 32000,
        'Prod_L': 14100,
        'Prod_M': 22600,
        'Prod_N': 11400,
        'Prod_O': 19800,
    }

    # Distribute ~26-27 transactions per product (400 / 15 ~ 26.67)
    # Give each product a count, totaling 400
    base_count = 400 // 15  # 26
    remainder = 400 - base_count * 15  # 10
    product_counts = {}
    for i, p in enumerate(products):
        product_counts[p] = base_count + (1 if i < remainder else 0)

    # Build transactions
    rows = []
    txn_id = 1
    start_date = datetime(2025, 1, 1)

    for product in products:
        count = product_counts[product]
        target = product_targets[product]

        # Distribute target revenue across 'count' transactions
        # Use random splits that sum to target exactly
        if count == 1:
            amounts = [target]
        else:
            # Generate count-1 random breakpoints, then compute differences
            avg = target / count
            amounts = []
            remaining = target
            for j in range(count - 1):
                # Random amount around average, ensuring we don't go negative
                lo = max(50, int(avg * 0.3))
                hi = min(int(avg * 2.5), remaining - (count - 1 - j) * 50)
                if hi <= lo:
                    hi = lo + 100
                amt = random.randint(lo, hi)
                amounts.append(amt)
                remaining -= amt
            amounts.append(remaining)  # last one gets the remainder

        for j, amt in enumerate(amounts):
            date = start_date + timedelta(days=random.randint(0, 364))
            region = random.choice(regions)
            rows.append((txn_id, date, product, region, amt))
            txn_id += 1

    # Shuffle rows to make it look realistic (not grouped by product)
    random.shuffle(rows)

    # Re-assign sequential TxnIDs after shuffle
    for i in range(len(rows)):
        rows[i] = (i + 1, rows[i][1], rows[i][2], rows[i][3], rows[i][4])

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ProductRevenue'

    # Headers
    headers = ['TxnID', 'Date', 'Product', 'Region', 'Revenue']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Data rows
    for r, row_data in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=row_data[0])  # TxnID
        ws.cell(row=r, column=2, value=row_data[1])  # Date
        ws.cell(row=r, column=2).number_format = 'yyyy-mm-dd'
        ws.cell(row=r, column=3, value=row_data[2])  # Product
        ws.cell(row=r, column=4, value=row_data[3])  # Region
        ws.cell(row=r, column=5, value=row_data[4])  # Revenue

    # Set column widths for readability
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Verify sums
    from collections import defaultdict
    sums = defaultdict(int)
    for row_data in rows:
        sums[row_data[2]] += row_data[4]
    print("Product revenue sums:")
    for p in sorted(sums.keys(), key=lambda x: sums[x], reverse=True):
        print(f"  {p}: {sums[p]}")

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
