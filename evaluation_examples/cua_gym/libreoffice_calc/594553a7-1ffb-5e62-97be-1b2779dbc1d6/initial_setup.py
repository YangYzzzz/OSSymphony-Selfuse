"""
Initial Setup: Paste Special Multiply operation on price data
Task ID: calc_cop_paste_special_008
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_cop_paste_special_008'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'PriceUpdate'

    # --- Headers ---
    headers = ['Product', 'Price', 'Category', 'Multiplier', 'Notes']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # --- Data rows: 19 products (rows 2-20) ---
    # Prices from context: B2=100, B3=250, B4=75; rest are realistic values
    # Multipliers from context: D2=1.05, D3=1.10, D4=1.08; rest are realistic
    data = [
        # Product,               Price,  Category,        Multiplier, Notes
        ('Laptop Pro 15"',       100,    'Electronics',   1.05,       'Annual price review'),
        ('Wireless Headphones',  250,    'Electronics',   1.10,       'Supply chain increase'),
        ('Office Chair Deluxe',  75,     'Furniture',     1.08,       'Material cost rise'),
        ('Mechanical Keyboard',  120,    'Electronics',   1.06,       'Component shortage'),
        ('Standing Desk',        320,    'Furniture',     1.12,       'Lumber surcharge'),
        ('USB-C Hub 7-Port',     45,     'Accessories',   1.07,       'Import tariff'),
        ('4K Monitor 27"',       480,    'Electronics',   1.09,       'Panel price increase'),
        ('Ergonomic Mouse',      65,     'Accessories',   1.05,       'Annual review'),
        ('Webcam 1080p',         85,     'Electronics',   1.08,       'Chip shortage'),
        ('Desk Lamp LED',        35,     'Office',        1.06,       'Energy costs'),
        ('Cable Management Kit', 28,     'Accessories',   1.04,       'Logistics'),
        ('Laptop Stand Alumin.', 55,     'Accessories',   1.07,       'Material costs'),
        ('External SSD 1TB',     130,    'Storage',       1.11,       'NAND pricing'),
        ('Wireless Charger 15W', 40,     'Electronics',   1.05,       'Annual review'),
        ('Printer Laser Color',  390,    'Electronics',   1.13,       'Cartridge supply'),
        ('Paper Shredder 12-Sh', 95,     'Office',        1.08,       'Motor costs'),
        ('Label Maker Pro',      60,     'Office',        1.06,       'Tape costs'),
        ('Document Scanner',     210,    'Electronics',   1.10,       'Sensor pricing'),
        ('Binding Machine',      75,     'Office',        1.07,       'Spring costs'),
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])   # Product
        ws.cell(row=r, column=2, value=row_data[1])   # Price (B column)
        ws.cell(row=r, column=3, value=row_data[2])   # Category
        ws.cell(row=r, column=4, value=row_data[3])   # Multiplier (D column)
        ws.cell(row=r, column=5, value=row_data[4])   # Notes

    # Column widths
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 22

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: PriceUpdate')
    print(f'  B2:B20 - current prices (B2=100, B3=250, B4=75, ...)')
    print(f'  D2:D20 - multipliers (D2=1.05, D3=1.10, D4=1.08, ...)')


create_initial()
