"""
Reward Script: Copy price multipliers in D2:D20 and use Paste Special 'Multiply'
  to scale existing prices in B2:B20 in-place.
Task ID: calc_cop_paste_special_008
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): At least 15 of 19 B-column prices are correctly multiplied
  Component 2 (0.3): All 19 B-column prices exactly match expected multiplied values
  Component 3 (0.2): D2:D20 multiplier values remain unchanged AND B column was modified
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_cop_paste_special_008'

# Known initial prices (B2:B20) and multipliers (D2:D20) used to compute expected golden values
INITIAL_PRICES = {
    2: 100, 3: 250, 4: 75, 5: 120, 6: 320, 7: 45, 8: 480,
    9: 65, 10: 85, 11: 35, 12: 28, 13: 55, 14: 130, 15: 40,
    16: 390, 17: 95, 18: 60, 19: 210, 20: 75
}
MULTIPLIERS = {
    2: 1.05, 3: 1.10, 4: 1.08, 5: 1.06, 6: 1.12, 7: 1.07, 8: 1.09,
    9: 1.05, 10: 1.08, 11: 1.06, 12: 1.04, 13: 1.07, 14: 1.11, 15: 1.05,
    16: 1.13, 17: 1.08, 18: 1.06, 19: 1.10, 20: 1.07
}

# Compute expected values: initial_price * multiplier
EXPECTED_PRICES = {row: round(INITIAL_PRICES[row] * MULTIPLIERS[row], 10) for row in range(2, 21)}


def count_correct_b_prices(ws):
    """Count how many of B2:B20 match expected multiplied values."""
    correct = 0
    for row in range(2, 21):
        actual = ws.cell(row=row, column=2).value
        expected = EXPECTED_PRICES[row]
        if actual is not None:
            try:
                if abs(float(actual) - expected) < 0.01:
                    correct += 1
            except (ValueError, TypeError):
                pass
    return correct


def count_wrong_b_prices(ws):
    """Count how many of B2:B20 do NOT match expected multiplied values (returns failures)."""
    failures = []
    for row in range(2, 21):
        actual = ws.cell(row=row, column=2).value
        expected = EXPECTED_PRICES[row]
        if actual is None:
            failures.append((row, None, expected))
        else:
            try:
                if abs(float(actual) - expected) >= 0.01:
                    failures.append((row, float(actual), expected))
            except (ValueError, TypeError):
                failures.append((row, actual, expected))
    return failures


def count_wrong_d_multipliers(ws):
    """Count how many of D2:D20 differ from expected multiplier values."""
    failures = []
    for row in range(2, 21):
        actual = ws.cell(row=row, column=4).value
        expected = MULTIPLIERS[row]
        if actual is None:
            failures.append((row, None, expected))
        else:
            try:
                if abs(float(actual) - expected) >= 0.001:
                    failures.append((row, float(actual), expected))
            except (ValueError, TypeError):
                failures.append((row, actual, expected))
    return failures


def count_b_values_matching_initial(ws):
    """Count how many B-column values still match the INITIAL (pre-multiply) values."""
    still_initial = 0
    for row in range(2, 21):
        actual = ws.cell(row=row, column=2).value
        initial = INITIAL_PRICES[row]
        if actual is not None:
            try:
                if abs(float(actual) - initial) < 0.01:
                    still_initial += 1
            except (ValueError, TypeError):
                pass
    return still_initial


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: sheet must exist
    if 'PriceUpdate' not in wb.sheetnames:
        print("CRITICAL: Sheet 'PriceUpdate' not found.")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['PriceUpdate']

    # Component 1: At least 15 of 19 B-column prices are correctly multiplied (0.5 points)
    # This checks that the Paste Special Multiply operation was applied to B2:B20.
    # On the initial file, B2:B20 contains original prices (not multiplied), so this FAILS.
    # On the golden file, B2:B20 contains multiplied prices, so this PASSES.
    try:
        correct_count = count_correct_b_prices(ws)
        wrong_list = count_wrong_b_prices(ws)

        if correct_count >= 15:
            print(f"PASS: Component 1 — {correct_count}/19 B-column prices correctly multiplied (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — only {correct_count}/19 B-column prices correctly multiplied (need >=15)")
            for row, actual, expected in wrong_list[:5]:  # Show first 5 failures
                print(f"  Row {row}: expected ~{expected}, got {actual}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All 19 B-column prices exactly match expected multiplied values (0.3 points)
    # This is a stricter check — all rows must be correct, not just 15.
    # On the initial file, no rows are multiplied, so this FAILS.
    # On the golden file, all rows are correctly multiplied, so this PASSES.
    try:
        failures = count_wrong_b_prices(ws)
        if len(failures) == 0:
            print(f"PASS: Component 2 — All 19 B-column prices exactly match expected multiplied values (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — {len(failures)} of 19 B-column prices did not match expected values")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: D2:D20 multiplier values remain unchanged AND B column was modified (0.2 points)
    # The task should only modify B-column prices; multipliers in D should stay the same.
    # This component is gated on B-column being modified (confirming paste special occurred).
    # On the initial file: B is NOT modified, so the gate check fails -> 0.0 pts
    # On the golden file: B IS modified AND D is unchanged -> 0.2 pts
    try:
        d_failures = count_wrong_d_multipliers(ws)
        still_initial_count = count_b_values_matching_initial(ws)
        b_was_modified = (still_initial_count < 19)  # At least one B value changed from initial

        if b_was_modified and len(d_failures) == 0:
            print(f"PASS: Component 3 — D2:D20 multipliers remain unchanged AND B column was modified (0.2 pts)")
            total_score += 0.2
        elif not b_was_modified:
            print(f"FAIL: Component 3 — B column was not modified (paste special multiply did not occur)")
        else:
            print(f"FAIL: Component 3 — {len(d_failures)} D2:D20 multiplier values were altered (should remain unchanged)")
            for row, actual, expected in d_failures[:5]:
                print(f"  D{row}: expected {expected}, got {actual}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
