"""
Initial Setup: P&L spreadsheet with monthly columns for 3 years (2022-2024)
Task ID: calc_gen_grouping_053
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_grouping_053'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

# Column mapping: columns B through AK for months
# B=2, C=3, ..., M=13 (2022 Jan-Dec)
# N=14, O=15, ..., Y=25 (2023 Jan-Dec)
# Z=26, AA=27, ..., AK=37 (2024 Jan-Dec)

MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
          'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

YEARS = [2022, 2023, 2024]

# Revenue lines (rows 3-8)
REVENUE_LINES = [
    'Product Sales - North America',
    'Product Sales - Europe',
    'Product Sales - Asia Pacific',
    'Service Revenue',
    'License Fees',
    'Other Revenue',
]

# COGS lines (rows 11-18)
COGS_LINES = [
    'Raw Materials',
    'Direct Labor',
    'Manufacturing Overhead',
    'Quality Control',
    'Packaging & Shipping',
    'Warranty Costs',
    'Contract Manufacturing',
    'Inventory Adjustments',
]

# Operating Expense lines (rows 21-32)
OPEX_LINES = [
    'Salaries & Benefits',
    'Rent & Facilities',
    'Marketing & Advertising',
    'Research & Development',
    'Information Technology',
    'Travel & Entertainment',
    'Professional Services',
    'Depreciation & Amortization',
    'Insurance',
    'Utilities',
    'Office Supplies',
    'Miscellaneous',
]

# Realistic monthly data seeding
import random
random.seed(42)

def make_monthly_data(base, variance=0.15):
    """Generate 36 months of realistic data around a base value."""
    vals = []
    trend = base
    for y in range(3):
        for m in range(12):
            seasonal = 1.0 + 0.1 * (1 if m in [2,3,4,8,9,10] else -0.05)
            growth = 1.0 + 0.03 * y / 12
            noise = 1.0 + random.uniform(-variance, variance)
            vals.append(round(trend * seasonal * growth * noise, 2))
    return vals


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'PandL'

    # --- Styles ---
    header_font = Font(name='Calibri', bold=True, size=11)
    header_fill_2022 = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_fill_2023 = PatternFill(start_color='FF70AD47', end_color='FF70AD47', fill_type='solid')
    header_fill_2024 = PatternFill(start_color='FFED7D31', end_color='FFED7D31', fill_type='solid')
    month_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    total_fill = PatternFill(start_color='FFFFC000', end_color='FFFFC000', fill_type='solid')
    total_font = Font(name='Calibri', bold=True, size=11)
    cat_fill = PatternFill(start_color='FFDAE3F3', end_color='FFDAE3F3', fill_type='solid')
    white_font = Font(name='Calibri', bold=True, size=11, color='FFFFFFFF')
    thin = Side(style='thin', color='000000')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    num_fmt = '#,##0.00'

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=False)
    left_align = Alignment(horizontal='left', vertical='center')

    # --- Row 1: Year headers ---
    ws.cell(row=1, column=1, value='Category')
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=1).alignment = center_align

    # 2022: B(2) - M(13)
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=13)
    ws.cell(row=1, column=2, value='2022')
    ws.cell(row=1, column=2).font = Font(name='Calibri', bold=True, size=12, color='FFFFFFFF')
    ws.cell(row=1, column=2).fill = header_fill_2022
    ws.cell(row=1, column=2).alignment = center_align

    # 2023: N(14) - Y(25)
    ws.merge_cells(start_row=1, start_column=14, end_row=1, end_column=25)
    ws.cell(row=1, column=14, value='2023')
    ws.cell(row=1, column=14).font = Font(name='Calibri', bold=True, size=12, color='FFFFFFFF')
    ws.cell(row=1, column=14).fill = header_fill_2023
    ws.cell(row=1, column=14).alignment = center_align

    # 2024: Z(26) - AK(37)
    ws.merge_cells(start_row=1, start_column=26, end_row=1, end_column=37)
    ws.cell(row=1, column=26, value='2024')
    ws.cell(row=1, column=26).font = Font(name='Calibri', bold=True, size=12, color='FFFFFFFF')
    ws.cell(row=1, column=26).fill = header_fill_2024
    ws.cell(row=1, column=26).alignment = center_align

    # --- Row 2: Month labels ---
    ws.cell(row=2, column=1, value='')
    for year_idx in range(3):
        for m_idx, month in enumerate(MONTHS):
            col = 2 + year_idx * 12 + m_idx
            ws.cell(row=2, column=col, value=month)
            ws.cell(row=2, column=col).font = Font(name='Calibri', bold=True, size=10)
            ws.cell(row=2, column=col).fill = month_fill
            ws.cell(row=2, column=col).alignment = center_align

    # --- Revenue rows (3-8) ---
    revenue_data = []
    for line in REVENUE_LINES:
        base = random.randint(300000, 800000)
        revenue_data.append(make_monthly_data(base, 0.12))

    for i, (line, vals) in enumerate(zip(REVENUE_LINES, revenue_data)):
        row = 3 + i
        ws.cell(row=row, column=1, value=line)
        ws.cell(row=row, column=1).font = Font(name='Calibri', size=10)
        ws.cell(row=row, column=1).alignment = left_align
        for col_idx, val in enumerate(vals):
            col = 2 + col_idx
            cell = ws.cell(row=row, column=col, value=val)
            cell.number_format = num_fmt

    # --- Row 9: Revenue Total (SUM formulas) ---
    ws.cell(row=9, column=1, value='Revenue Total')
    ws.cell(row=9, column=1).font = total_font
    ws.cell(row=9, column=1).fill = total_fill
    ws.cell(row=9, column=1).alignment = left_align
    for col in range(2, 38):
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(col)
        cell = ws.cell(row=9, column=col, value=f'=SUM({col_letter}3:{col_letter}8)')
        cell.font = total_font
        cell.fill = total_fill
        cell.number_format = num_fmt

    # --- Row 10: Blank separator ---
    ws.cell(row=10, column=1, value='')

    # --- COGS rows (11-18) ---
    cogs_data = []
    for line in COGS_LINES:
        base = random.randint(80000, 250000)
        cogs_data.append(make_monthly_data(base, 0.10))

    for i, (line, vals) in enumerate(zip(COGS_LINES, cogs_data)):
        row = 11 + i
        ws.cell(row=row, column=1, value=line)
        ws.cell(row=row, column=1).font = Font(name='Calibri', size=10)
        ws.cell(row=row, column=1).alignment = left_align
        for col_idx, val in enumerate(vals):
            col = 2 + col_idx
            cell = ws.cell(row=row, column=col, value=val)
            cell.number_format = num_fmt

    # --- Row 19: COGS Total (SUM formulas) ---
    from openpyxl.utils import get_column_letter
    ws.cell(row=19, column=1, value='COGS Total')
    ws.cell(row=19, column=1).font = total_font
    ws.cell(row=19, column=1).fill = total_fill
    ws.cell(row=19, column=1).alignment = left_align
    for col in range(2, 38):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=19, column=col, value=f'=SUM({col_letter}11:{col_letter}18)')
        cell.font = total_font
        cell.fill = total_fill
        cell.number_format = num_fmt

    # --- Row 20: Blank separator ---
    ws.cell(row=20, column=1, value='')

    # --- Operating Expense rows (21-32) ---
    opex_data = []
    for line in OPEX_LINES:
        base = random.randint(40000, 180000)
        opex_data.append(make_monthly_data(base, 0.10))

    for i, (line, vals) in enumerate(zip(OPEX_LINES, opex_data)):
        row = 21 + i
        ws.cell(row=row, column=1, value=line)
        ws.cell(row=row, column=1).font = Font(name='Calibri', size=10)
        ws.cell(row=row, column=1).alignment = left_align
        for col_idx, val in enumerate(vals):
            col = 2 + col_idx
            cell = ws.cell(row=row, column=col, value=val)
            cell.number_format = num_fmt

    # --- Row 33: OpEx Total (SUM formulas) ---
    ws.cell(row=33, column=1, value='OpEx Total')
    ws.cell(row=33, column=1).font = total_font
    ws.cell(row=33, column=1).fill = total_fill
    ws.cell(row=33, column=1).alignment = left_align
    for col in range(2, 38):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=33, column=col, value=f'=SUM({col_letter}21:{col_letter}32)')
        cell.font = total_font
        cell.fill = total_fill
        cell.number_format = num_fmt

    # --- Row 34: Blank separator ---
    ws.cell(row=34, column=1, value='')

    # --- Row 35: Net Income ---
    ws.cell(row=35, column=1, value='Net Income')
    ws.cell(row=35, column=1).font = Font(name='Calibri', bold=True, size=11)
    ws.cell(row=35, column=1).alignment = left_align
    net_fill = PatternFill(start_color='FF375623', end_color='FF375623', fill_type='solid')
    net_font = Font(name='Calibri', bold=True, size=11, color='FFFFFFFF')
    ws.cell(row=35, column=1).fill = net_fill
    ws.cell(row=35, column=1).font = net_font
    for col in range(2, 38):
        col_letter = get_column_letter(col)
        # Net Income = Revenue - COGS - OpEx
        cell = ws.cell(row=35, column=col,
                       value=f'={col_letter}9-{col_letter}19-{col_letter}33')
        cell.font = net_font
        cell.fill = net_fill
        cell.number_format = num_fmt

    # --- Column widths ---
    ws.column_dimensions['A'].width = 32
    for col in range(2, 38):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 12

    # --- Row heights ---
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[9].height = 18
    ws.row_dimensions[19].height = 18
    ws.row_dimensions[33].height = 18
    ws.row_dimensions[35].height = 22

    # Freeze panes at B3
    ws.freeze_panes = 'B3'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
