"""
Reward Script: P&L Spreadsheet Row/Column Grouping with SUBTOTAL formulas
Task ID: calc_gen_grouping_053
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Column grouping (B-M, N-Y, Z-AK all at outline_level=1)   — 0.35 pts
  Component 2: Row grouping (rows 3-8, 11-18, 21-32 all at outline_level=1) — 0.35 pts
  Component 3: SUBTOTAL formulas in Revenue Total row 9                    — 0.10 pts
  Component 4: SUBTOTAL formulas in COGS Total row 19                      — 0.10 pts
  Component 5: SUBTOTAL formulas in OpEx Total row 33                      — 0.10 pts
  Total: 1.0
"""

import os
import openpyxl
from openpyxl.utils import column_index_from_string

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_gen_grouping_053'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: PandL sheet must exist
    if 'PandL' not in wb.sheetnames:
        print("CRITICAL: Sheet 'PandL' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['PandL']

    # --- Component 1: Column grouping (0.35 points) ---
    # Task requires B-M (2022 months), N-Y (2023 months), Z-AK (2024 months) all grouped at level 1
    # This FAILS on initial (no column groupings) → PASSES on golden (all month cols grouped)
    try:
        # Expected: columns B through AK (indices 2-37) should all be at outline_level=1
        # Column A is the label column and should NOT be grouped
        expected_grouped_cols = set()
        for idx in range(2, 38):  # B=2 through AK=37
            from openpyxl.utils import get_column_letter
            expected_grouped_cols.add(get_column_letter(idx))

        grouped_cols = set()
        for col_name, cd in ws.column_dimensions.items():
            if cd.outline_level >= 1:
                grouped_cols.add(col_name)

        # Check that all expected columns are grouped at outline_level=1
        b_to_m_grouped = all(
            ws.column_dimensions.get(col, None) is not None and
            ws.column_dimensions[col].outline_level >= 1
            for col in [get_column_letter(i) for i in range(2, 14)]  # B-M
        )
        n_to_y_grouped = all(
            ws.column_dimensions.get(col, None) is not None and
            ws.column_dimensions[col].outline_level >= 1
            for col in [get_column_letter(i) for i in range(14, 26)]  # N-Y
        )
        z_to_ak_grouped = all(
            ws.column_dimensions.get(col, None) is not None and
            ws.column_dimensions[col].outline_level >= 1
            for col in [get_column_letter(i) for i in range(26, 38)]  # Z-AK
        )

        if b_to_m_grouped and n_to_y_grouped and z_to_ak_grouped:
            print("PASS: Component 1 — All month columns B-M, N-Y, Z-AK grouped at outline_level=1 (0.35 pts)")
            total_score += 0.35
        else:
            parts = []
            if not b_to_m_grouped:
                parts.append("B-M (2022 months) not fully grouped")
            if not n_to_y_grouped:
                parts.append("N-Y (2023 months) not fully grouped")
            if not z_to_ak_grouped:
                parts.append("Z-AK (2024 months) not fully grouped")
            print(f"FAIL: Component 1 — Column grouping incomplete: {'; '.join(parts)}")
    except Exception as e:
        print(f"ERROR: Component 1 (column grouping) — {e}")

    # --- Component 2: Row grouping (0.35 points) ---
    # Task requires rows 3-8 (Revenue detail), 11-18 (COGS detail), 21-32 (OpEx detail) grouped at level 1
    # This FAILS on initial (no row groupings) → PASSES on golden
    try:
        revenue_rows_grouped = all(
            ws.row_dimensions.get(r, None) is not None and
            ws.row_dimensions[r].outline_level >= 1
            for r in range(3, 9)  # rows 3-8
        )
        cogs_rows_grouped = all(
            ws.row_dimensions.get(r, None) is not None and
            ws.row_dimensions[r].outline_level >= 1
            for r in range(11, 19)  # rows 11-18
        )
        opex_rows_grouped = all(
            ws.row_dimensions.get(r, None) is not None and
            ws.row_dimensions[r].outline_level >= 1
            for r in range(21, 33)  # rows 21-32
        )

        if revenue_rows_grouped and cogs_rows_grouped and opex_rows_grouped:
            print("PASS: Component 2 — All expense rows 3-8, 11-18, 21-32 grouped at outline_level=1 (0.35 pts)")
            total_score += 0.35
        else:
            parts = []
            if not revenue_rows_grouped:
                parts.append("rows 3-8 (Revenue detail) not fully grouped")
            if not cogs_rows_grouped:
                parts.append("rows 11-18 (COGS detail) not fully grouped")
            if not opex_rows_grouped:
                parts.append("rows 21-32 (OpEx detail) not fully grouped")
            print(f"FAIL: Component 2 — Row grouping incomplete: {'; '.join(parts)}")
    except Exception as e:
        print(f"ERROR: Component 2 (row grouping) — {e}")

    # --- Component 3: SUBTOTAL formulas in Revenue Total row 9 (0.10 points) ---
    # Task requires replacing SUM with SUBTOTAL(9,...) in row 9 so collapsed rows are excluded
    # This FAILS on initial (still SUM) → PASSES on golden (SUBTOTAL)
    try:
        row9_subtotal_count = 0
        row9_total_formula_cells = 0
        for col in range(2, 38):  # B through AK
            val = ws.cell(row=9, column=col).value
            if val is not None and isinstance(val, str) and val.strip():
                row9_total_formula_cells += 1
                # Check for SUBTOTAL(9,...) pattern — case-insensitive
                val_upper = val.upper().replace(' ', '')
                if 'SUBTOTAL(9,' in val_upper or 'SUBTOTAL(9,' in val_upper:
                    row9_subtotal_count += 1

        if row9_total_formula_cells > 0 and row9_subtotal_count == row9_total_formula_cells:
            print(f"PASS: Component 3 — All {row9_subtotal_count} formula cells in Revenue Total row 9 use SUBTOTAL(9,...) (0.10 pts)")
            total_score += 0.10
        elif row9_total_formula_cells == 0:
            print("FAIL: Component 3 — No formula cells found in row 9")
        else:
            print(f"FAIL: Component 3 — Only {row9_subtotal_count}/{row9_total_formula_cells} cells in row 9 use SUBTOTAL(9,...)")
    except Exception as e:
        print(f"ERROR: Component 3 (row 9 SUBTOTAL) — {e}")

    # --- Component 4: SUBTOTAL formulas in COGS Total row 19 (0.10 points) ---
    # Task requires replacing SUM with SUBTOTAL(9,...) in row 19
    # This FAILS on initial (still SUM) → PASSES on golden (SUBTOTAL)
    try:
        row19_subtotal_count = 0
        row19_total_formula_cells = 0
        for col in range(2, 38):  # B through AK
            val = ws.cell(row=19, column=col).value
            if val is not None and isinstance(val, str) and val.strip():
                row19_total_formula_cells += 1
                val_upper = val.upper().replace(' ', '')
                if 'SUBTOTAL(9,' in val_upper:
                    row19_subtotal_count += 1

        if row19_total_formula_cells > 0 and row19_subtotal_count == row19_total_formula_cells:
            print(f"PASS: Component 4 — All {row19_subtotal_count} formula cells in COGS Total row 19 use SUBTOTAL(9,...) (0.10 pts)")
            total_score += 0.10
        elif row19_total_formula_cells == 0:
            print("FAIL: Component 4 — No formula cells found in row 19")
        else:
            print(f"FAIL: Component 4 — Only {row19_subtotal_count}/{row19_total_formula_cells} cells in row 19 use SUBTOTAL(9,...)")
    except Exception as e:
        print(f"ERROR: Component 4 (row 19 SUBTOTAL) — {e}")

    # --- Component 5: SUBTOTAL formulas in OpEx Total row 33 (0.10 points) ---
    # Task requires replacing SUM with SUBTOTAL(9,...) in row 33
    # This FAILS on initial (still SUM) → PASSES on golden (SUBTOTAL)
    try:
        row33_subtotal_count = 0
        row33_total_formula_cells = 0
        for col in range(2, 38):  # B through AK
            val = ws.cell(row=33, column=col).value
            if val is not None and isinstance(val, str) and val.strip():
                row33_total_formula_cells += 1
                val_upper = val.upper().replace(' ', '')
                if 'SUBTOTAL(9,' in val_upper:
                    row33_subtotal_count += 1

        if row33_total_formula_cells > 0 and row33_subtotal_count == row33_total_formula_cells:
            print(f"PASS: Component 5 — All {row33_subtotal_count} formula cells in OpEx Total row 33 use SUBTOTAL(9,...) (0.10 pts)")
            total_score += 0.10
        elif row33_total_formula_cells == 0:
            print("FAIL: Component 5 — No formula cells found in row 33")
        else:
            print(f"FAIL: Component 5 — Only {row33_subtotal_count}/{row33_total_formula_cells} cells in row 33 use SUBTOTAL(9,...)")
    except Exception as e:
        print(f"ERROR: Component 5 (row 33 SUBTOTAL) — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
