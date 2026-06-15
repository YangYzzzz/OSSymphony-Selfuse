"""
Reward Script: Fix broken TaxRate named range
Task ID: calc_tbl_059
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): TaxRate named range exists and does NOT reference #REF
  Component 2 (0.5): TaxRate named range points to Sheet1!$G$1 specifically
  Component 3 (0.2): Named range target cell contains correct tax rate 0.0825
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_059'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    The task requires fixing the named range 'TaxRate' which points to a deleted
    sheet (#REF!$F$1) so that it points to Sheet1!$G$1 where the tax rate 0.0825
    is stored. All formulas using TaxRate (E2:E13) then calculate correctly.

    Key difference between initial and golden:
      initial: TaxRate -> '#REF'!$F$1  (broken)
      golden:  TaxRate -> Sheet1!$G$1  (fixed)
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: TaxRate named range exists and does NOT reference #REF (0.3 points)
    # Initial: TaxRate -> '#REF'!$F$1 (contains #REF -> FAIL)
    # Golden:  TaxRate -> Sheet1!$G$1 (no #REF -> PASS)
    try:
        defined_names = dict(wb.defined_names)
        if 'TaxRate' in defined_names:
            attr_text = defined_names['TaxRate'].attr_text
            if '#REF' in attr_text.upper():
                print(f"FAIL: Component 1 -- TaxRate still references #REF: {attr_text}")
            else:
                print(f"PASS: Component 1 -- TaxRate exists and has no #REF (value: {attr_text}) (0.3 pts)")
                total_score += 0.3
        else:
            print(f"FAIL: Component 1 -- TaxRate named range not found. Names: {list(defined_names.keys())}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: TaxRate named range points to Sheet1!$G$1 specifically (0.5 points)
    # Initial: '#REF'!$F$1 (FAIL - wrong sheet and wrong cell)
    # Golden:  Sheet1!$G$1 (PASS)
    try:
        defined_names = dict(wb.defined_names)
        if 'TaxRate' in defined_names:
            attr_text = defined_names['TaxRate'].attr_text
            # Normalize: remove quotes, compare case-insensitively
            normalized = attr_text.replace("'", "").replace('"', '').upper().replace(' ', '')
            expected = "SHEET1!$G$1"
            if normalized == expected:
                print(f"PASS: Component 2 -- TaxRate points to Sheet1!$G$1 (raw: {attr_text}) (0.5 pts)")
                total_score += 0.5
            else:
                print(f"FAIL: Component 2 -- TaxRate points to {attr_text}, expected Sheet1!$G$1")
        else:
            print(f"FAIL: Component 2 -- TaxRate named range not found")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: The target cell (Sheet1 G1) contains the tax rate 0.0825 AND
    # the named range references that cell (combined check) (0.2 points)
    # Initial: Named range points to #REF (can't resolve) -> FAIL
    # Golden:  Named range points to Sheet1!$G$1 which has 0.0825 -> PASS
    try:
        defined_names = dict(wb.defined_names)
        if 'TaxRate' in defined_names:
            attr_text = defined_names['TaxRate'].attr_text
            # Parse the reference to extract sheet and cell
            # Expected format: Sheet1!$G$1 or 'Sheet1'!$G$1
            match = re.match(r"'?([^'!]+)'?!\$?([A-Z]+)\$?(\d+)", attr_text)
            if match:
                sheet_name = match.group(1)
                col_letter = match.group(2)
                row_num = int(match.group(3))
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    cell_val = ws[f"{col_letter}{row_num}"].value
                    if cell_val is not None and abs(float(cell_val) - 0.0825) < 0.0001:
                        print(f"PASS: Component 3 -- Named range target {sheet_name}!{col_letter}{row_num} contains {cell_val} (expected 0.0825) (0.2 pts)")
                        total_score += 0.2
                    else:
                        print(f"FAIL: Component 3 -- Target cell value is {cell_val}, expected 0.0825")
                else:
                    print(f"FAIL: Component 3 -- Sheet '{sheet_name}' not found in workbook")
            else:
                print(f"FAIL: Component 3 -- Cannot parse named range reference: {attr_text}")
        else:
            print(f"FAIL: Component 3 -- TaxRate named range not found")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
