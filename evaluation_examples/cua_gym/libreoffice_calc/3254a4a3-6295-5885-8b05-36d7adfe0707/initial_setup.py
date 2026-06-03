"""
Initial Setup: Named range 'TaxRate' points to deleted sheet causing #REF! errors
Task ID: calc_tbl_059
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.workbook.defined_name import DefinedName

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_059'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: Invoice Items ---
    ws = wb.active
    ws.title = "Sheet1"

    # Header row styling
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # Headers
    headers = ["Item", "Unit Price", "Quantity", "Subtotal", "Tax", "Total"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # G1 holds the tax rate value
    ws.cell(row=1, column=7, value=0.0825)
    ws["G1"].font = Font(name="Arial", size=11, bold=True)
    ws["G1"].number_format = '0.0000'

    # Label for G1
    ws.cell(row=2, column=7, value="Tax Rate")
    ws["G2"].font = Font(name="Arial", size=9, italic=True, color="666666")

    # Invoice data rows
    items = [
        ["Wireless Keyboard", 49.99, 12],
        ["USB-C Hub Adapter", 34.95, 25],
        ["Ergonomic Mouse", 29.99, 18],
        ["Monitor Stand", 89.50, 7],
        ["Laptop Sleeve 15\"", 24.99, 30],
        ["Webcam HD 1080p", 59.99, 15],
        ["Desk Lamp LED", 42.50, 10],
        ["Cable Management Kit", 19.95, 22],
        ["Screen Protector Pack", 12.99, 40],
        ["Bluetooth Speaker", 75.00, 8],
        ["Phone Stand", 15.99, 35],
        ["HDMI Cable 6ft", 9.99, 50],
    ]

    currency_fmt = '$#,##0.00'
    data_align = Alignment(horizontal="center", vertical="center")

    for r, (item, price, qty) in enumerate(items, 2):
        ws.cell(row=r, column=1, value=item).border = thin_border
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")

        c_price = ws.cell(row=r, column=2, value=price)
        c_price.number_format = currency_fmt
        c_price.border = thin_border
        c_price.alignment = data_align

        c_qty = ws.cell(row=r, column=3, value=qty)
        c_qty.number_format = '0'
        c_qty.border = thin_border
        c_qty.alignment = data_align

        # Subtotal = Unit Price * Quantity
        c_sub = ws.cell(row=r, column=4, value=f'=B{r}*C{r}')
        c_sub.number_format = currency_fmt
        c_sub.border = thin_border
        c_sub.alignment = data_align

        # Tax = Subtotal * TaxRate  (will show #REF! because TaxRate points to deleted sheet)
        c_tax = ws.cell(row=r, column=5, value=f'=D{r}*TaxRate')
        c_tax.number_format = currency_fmt
        c_tax.border = thin_border
        c_tax.alignment = data_align

        # Total = Subtotal + Tax
        c_total = ws.cell(row=r, column=6, value=f'=D{r}+E{r}')
        c_total.number_format = currency_fmt
        c_total.border = thin_border
        c_total.alignment = data_align

    # Summary row
    last_row = len(items) + 1
    summary_row = last_row + 1
    ws.cell(row=summary_row, column=3, value="TOTALS:").font = Font(bold=True, size=11)
    ws.cell(row=summary_row, column=3).alignment = Alignment(horizontal="right")

    ws.cell(row=summary_row, column=4, value=f'=SUM(D2:D{last_row})')
    ws.cell(row=summary_row, column=4).number_format = currency_fmt
    ws.cell(row=summary_row, column=4).font = Font(bold=True)
    ws.cell(row=summary_row, column=4).border = Border(top=Side(style="double", color="000000"), bottom=Side(style="double", color="000000"))

    ws.cell(row=summary_row, column=5, value=f'=SUM(E2:E{last_row})')
    ws.cell(row=summary_row, column=5).number_format = currency_fmt
    ws.cell(row=summary_row, column=5).font = Font(bold=True)
    ws.cell(row=summary_row, column=5).border = Border(top=Side(style="double", color="000000"), bottom=Side(style="double", color="000000"))

    ws.cell(row=summary_row, column=6, value=f'=SUM(F2:F{last_row})')
    ws.cell(row=summary_row, column=6).number_format = currency_fmt
    ws.cell(row=summary_row, column=6).font = Font(bold=True)
    ws.cell(row=summary_row, column=6).border = Border(top=Side(style="double", color="000000"), bottom=Side(style="double", color="000000"))

    # Column widths
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 12

    # Freeze header row
    ws.freeze_panes = "A2"

    # --- Define the BROKEN named range 'TaxRate' pointing to a deleted sheet ---
    # We reference a non-existent sheet to simulate #REF! error
    # In openpyxl, we can define a named range with an invalid reference
    broken_ref = "'#REF'!$F$1"
    dn = DefinedName(name='TaxRate', attr_text=broken_ref)
    wb.defined_names.add(dn)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
