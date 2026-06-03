"""
Reward Script: Count total number of orders per warehouse location in Sheet2
Task ID: osworld_calc_pivot_count_invoice_011
Domain: libreoffice_calc
Scoring:
  Component 1: Sheet2 has a header row with location and count columns (0.2 pts)
  Component 2: Sheet2 has exactly 5 data rows (one per unique warehouse) (0.3 pts)
  Component 3: All 5 warehouse locations are present in Sheet2 (0.2 pts)
  Component 4: Each warehouse location has the correct order count (0.3 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_pivot_count_invoice_011'

# Ground truth: expected counts per warehouse location derived from Sheet1 data
EXPECTED_COUNTS = {
    'Atlanta': 4,
    'Chicago': 5,
    'Dallas': 4,
    'New York': 3,
    'Seattle': 4,
}


def verify_task(file_path):
    """
    Verify that Sheet2 contains a pivot-style summary of order counts per
    warehouse location, matching the data in Sheet1.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sheet2 must exist
    if 'Sheet2' not in wb.sheetnames:
        print("CRITICAL: Sheet2 not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws2 = wb['Sheet2']

    # -----------------------------------------------------------------------
    # Component 1: Sheet2 has a header row with a location column and a count
    # column (0.2 points).
    # On initial_env, Sheet2 is empty — this check will FAIL there.
    # -----------------------------------------------------------------------
    try:
        header_row = [ws2.cell(row=1, column=c).value for c in range(1, ws2.max_column + 1)]
        # Look for at least one cell referencing "warehouse" / "location" (case-insensitive)
        # and one cell referencing "count" / "orders"
        has_location_header = any(
            v and ('location' in str(v).lower() or 'warehouse' in str(v).lower())
            for v in header_row
        )
        has_count_header = any(
            v and ('count' in str(v).lower() or 'order' in str(v).lower() or 'total' in str(v).lower())
            for v in header_row
        )
        if has_location_header and has_count_header:
            print(f"PASS: Component 1 — header row found with location and count columns: {header_row} (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 1 — expected header row with location+count columns, found: {header_row}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Sheet2 has exactly 5 data rows (one per unique warehouse
    # location from Sheet1) (0.3 points).
    # On initial_env, Sheet2 is empty — this check will FAIL there.
    # -----------------------------------------------------------------------
    try:
        # Count non-empty rows beyond the header
        data_rows = []
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                data_rows.append(row)

        expected_row_count = len(EXPECTED_COUNTS)  # 5
        if len(data_rows) == expected_row_count:
            print(f"PASS: Component 2 — Sheet2 has exactly {expected_row_count} data rows (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — expected {expected_row_count} data rows, found {len(data_rows)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: All 5 unique warehouse locations are present in Sheet2
    # (0.2 points).
    # On initial_env, Sheet2 is empty — this check will FAIL there.
    # -----------------------------------------------------------------------
    try:
        # Find the column that contains warehouse location names
        location_col = None
        for c in range(1, ws2.max_column + 1):
            header_val = ws2.cell(row=1, column=c).value
            if header_val and ('location' in str(header_val).lower() or 'warehouse' in str(header_val).lower()):
                location_col = c
                break

        if location_col is None:
            # Fallback: assume first column is the location column
            location_col = 1

        found_locations = set()
        for row in ws2.iter_rows(min_row=2, values_only=False):
            loc_val = ws2.cell(row=row[0].row, column=location_col).value
            if loc_val is not None:
                found_locations.add(str(loc_val).strip())

        expected_locations = set(EXPECTED_COUNTS.keys())
        if found_locations == expected_locations:
            print(f"PASS: Component 3 — all 5 warehouse locations found: {sorted(found_locations)} (0.2 pts)")
            total_score += 0.2
        else:
            missing = expected_locations - found_locations
            extra = found_locations - expected_locations
            print(f"FAIL: Component 3 — locations mismatch. Missing: {missing}, Extra: {extra}, Found: {sorted(found_locations)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: Each warehouse location has the correct order count
    # (0.3 points).
    # On initial_env, Sheet2 is empty — this check will FAIL there.
    # -----------------------------------------------------------------------
    try:
        # Find location column and count column
        loc_col = None
        cnt_col = None
        for c in range(1, ws2.max_column + 1):
            header_val = ws2.cell(row=1, column=c).value
            if header_val:
                hv_lower = str(header_val).lower()
                if 'location' in hv_lower or 'warehouse' in hv_lower:
                    loc_col = c
                elif 'count' in hv_lower or 'order' in hv_lower or 'total' in hv_lower:
                    cnt_col = c

        if loc_col is None:
            loc_col = 1
        if cnt_col is None:
            cnt_col = 2

        # Build actual counts from Sheet2
        actual_counts = {}
        for r in range(2, ws2.max_row + 1):
            loc_val = ws2.cell(row=r, column=loc_col).value
            cnt_val = ws2.cell(row=r, column=cnt_col).value
            if loc_val is not None and cnt_val is not None:
                try:
                    actual_counts[str(loc_val).strip()] = int(cnt_val)
                except (ValueError, TypeError):
                    actual_counts[str(loc_val).strip()] = cnt_val

        # Compare with expected — count correct matches
        mismatches = []
        for location, expected_count in EXPECTED_COUNTS.items():
            actual_count = actual_counts.get(location)
            if actual_count != expected_count:
                mismatches.append(f"{location}: expected {expected_count}, found {actual_count}")

        if len(mismatches) == 0 and len(actual_counts) == len(EXPECTED_COUNTS):
            print(f"PASS: Component 4 — all warehouse counts correct: {actual_counts} (0.3 pts)")
            total_score += 0.3
        elif len(mismatches) > 0:
            print(f"FAIL: Component 4 — count mismatches: {'; '.join(mismatches)}")
        else:
            print(f"FAIL: Component 4 — wrong number of count entries: {actual_counts}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the VM environment
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
