"""
Initial Setup: Order fulfillment data with empty Sheet2 (no pivot table)
Task ID: osworld_calc_pivot_count_invoice_011
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_pivot_count_invoice_011'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: Order Fulfillment Data ---
    ws1 = wb.active
    ws1.title = 'Sheet1'

    # Headers
    headers = ['Order ID', 'Warehouse Location', 'Product SKU', 'Order Date', 'Fulfillment Status']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # Realistic order fulfillment data across multiple warehouse locations
    data = [
        ['ORD-10001', 'Chicago',      'SKU-A204', '2025-01-03', 'Shipped'],
        ['ORD-10002', 'Dallas',        'SKU-B117', '2025-01-04', 'Delivered'],
        ['ORD-10003', 'Chicago',      'SKU-C389', '2025-01-05', 'Delivered'],
        ['ORD-10004', 'Seattle',       'SKU-A204', '2025-01-06', 'Processing'],
        ['ORD-10005', 'Dallas',        'SKU-D501', '2025-01-07', 'Shipped'],
        ['ORD-10006', 'Atlanta',       'SKU-B117', '2025-01-08', 'Delivered'],
        ['ORD-10007', 'Chicago',      'SKU-E672', '2025-01-09', 'Delivered'],
        ['ORD-10008', 'Seattle',       'SKU-C389', '2025-01-10', 'Shipped'],
        ['ORD-10009', 'Atlanta',       'SKU-A204', '2025-01-11', 'Delivered'],
        ['ORD-10010', 'Dallas',        'SKU-F830', '2025-01-12', 'Processing'],
        ['ORD-10011', 'Chicago',      'SKU-B117', '2025-01-13', 'Shipped'],
        ['ORD-10012', 'New York',      'SKU-D501', '2025-01-14', 'Delivered'],
        ['ORD-10013', 'Seattle',       'SKU-A204', '2025-01-15', 'Delivered'],
        ['ORD-10014', 'New York',      'SKU-E672', '2025-01-16', 'Shipped'],
        ['ORD-10015', 'Atlanta',       'SKU-C389', '2025-01-17', 'Delivered'],
        ['ORD-10016', 'Dallas',        'SKU-B117', '2025-01-18', 'Delivered'],
        ['ORD-10017', 'Chicago',      'SKU-F830', '2025-01-19', 'Shipped'],
        ['ORD-10018', 'New York',      'SKU-A204', '2025-01-20', 'Processing'],
        ['ORD-10019', 'Seattle',       'SKU-D501', '2025-01-21', 'Delivered'],
        ['ORD-10020', 'Atlanta',       'SKU-E672', '2025-01-22', 'Shipped'],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # --- Sheet2: Empty (agent needs to create pivot table here) ---
    ws2 = wb.create_sheet('Sheet2')
    # Sheet2 intentionally left empty — task asks agent to create pivot count here

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
