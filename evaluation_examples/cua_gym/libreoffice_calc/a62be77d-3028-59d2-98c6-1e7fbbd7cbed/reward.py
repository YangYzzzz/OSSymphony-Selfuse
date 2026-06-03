"""
Reward Script: Book Club Reading List Tracker
Task ID: calc_grs_038
Domain: libreoffice_calc
Scoring:
  Component 1: Days to Read formulas in column I (0.15)
  Component 2: Average Member Rating formulas in column P (0.15)
  Component 3: Data validations for Genre and Status dropdowns (0.15)
  Component 4: Conditional formatting for rating >= 4 (0.10)
  Component 5: Summary sheet populated with formulas (0.20)
  Component 6: Bar chart on Summary sheet (0.15)
  Component 7: Sort order — Date Finished descending (0.10)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_038'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Reading List' sheet must exist
    if 'Reading List' not in wb.sheetnames:
        print("CRITICAL: 'Reading List' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Reading List']

    # ---------------------------------------------------------------
    # Component 1: Days to Read formulas in column I (0.15 points)
    # Task requires: Days to Read (formula) — should be =H-G for each completed book
    # Initial has NO formulas in column I; golden has =H<row>-G<row>
    # ---------------------------------------------------------------
    try:
        formula_count = 0
        expected_rows = 0
        for r in range(2, ws.max_row + 1):
            date_started = ws.cell(row=r, column=7).value   # G
            date_finished = ws.cell(row=r, column=8).value   # H
            # Only rows with both dates should have a formula
            if date_started is not None and date_finished is not None:
                expected_rows += 1
                cell_val = ws.cell(row=r, column=9).value  # I
                if isinstance(cell_val, str) and '=' in cell_val:
                    formula_count += 1

        if expected_rows > 0 and formula_count >= expected_rows:
            print(f"PASS: Component 1 — Days to Read formulas found ({formula_count}/{expected_rows} rows) (0.15 pts)")
            total_score += 0.15
        elif expected_rows > 0 and formula_count > 0:
            partial = 0.15 * (formula_count / expected_rows)
            print(f"PARTIAL: Component 1 — Days to Read formulas {formula_count}/{expected_rows} rows ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No Days to Read formulas found (expected {expected_rows})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ---------------------------------------------------------------
    # Component 2: Average Member Rating formulas in column P (0.15 points)
    # Task requires: Average Member Rating (formula) — should be =AVERAGE(K:O) for completed books
    # Initial has NO formulas in column P; golden has =AVERAGE(K<row>:O<row>)
    # ---------------------------------------------------------------
    try:
        avg_formula_count = 0
        expected_avg_rows = 0
        for r in range(2, ws.max_row + 1):
            # Check if row has member reviews (K-O columns)
            has_reviews = any(ws.cell(row=r, column=c).value is not None for c in range(11, 16))
            if has_reviews:
                expected_avg_rows += 1
                cell_val = ws.cell(row=r, column=16).value  # P
                if isinstance(cell_val, str) and 'AVERAGE' in cell_val.upper():
                    avg_formula_count += 1

        if expected_avg_rows > 0 and avg_formula_count >= expected_avg_rows:
            print(f"PASS: Component 2 — Average Member Rating formulas found ({avg_formula_count}/{expected_avg_rows}) (0.15 pts)")
            total_score += 0.15
        elif expected_avg_rows > 0 and avg_formula_count > 0:
            partial = 0.15 * (avg_formula_count / expected_avg_rows)
            print(f"PARTIAL: Component 2 — Average formulas {avg_formula_count}/{expected_avg_rows} ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No Average Member Rating formulas found (expected {expected_avg_rows})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---------------------------------------------------------------
    # Component 3: Data validations for Genre and Status (0.15 points)
    # Task requires: Genre dropdown (Fiction, Non-Fiction, Mystery, Sci-Fi, Biography, Self-Help, Fantasy)
    #                Status dropdown (To Read, Currently Reading, Completed, Abandoned)
    # Initial has 0 data validations; golden has 2
    # ---------------------------------------------------------------
    try:
        dvs = ws.data_validations.dataValidation if ws.data_validations else []
        genre_dv_found = False
        status_dv_found = False

        for dv in dvs:
            if dv.type == 'list' and dv.formula1:
                formula_lower = dv.formula1.lower()
                # Check for genre dropdown
                if 'fiction' in formula_lower and 'mystery' in formula_lower:
                    genre_dv_found = True
                # Check for status dropdown
                if 'completed' in formula_lower and 'reading' in formula_lower:
                    status_dv_found = True

        if genre_dv_found and status_dv_found:
            print(f"PASS: Component 3 — Both Genre and Status data validations found (0.15 pts)")
            total_score += 0.15
        elif genre_dv_found or status_dv_found:
            found = "Genre" if genre_dv_found else "Status"
            print(f"PARTIAL: Component 3 — Only {found} validation found (0.075 pts)")
            total_score += 0.075
        else:
            print(f"FAIL: Component 3 — No Genre/Status data validations found (have {len(dvs)} validations)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ---------------------------------------------------------------
    # Component 4: Conditional formatting for rating >= 4 (0.10 points)
    # Task requires: gold background for books rated 4+ stars
    # Initial has 0 conditional formatting rules; golden has expression rule $J2>=4
    # ---------------------------------------------------------------
    try:
        cfs = list(ws.conditional_formatting)
        rating_cf_found = False

        for cf in cfs:
            for rule in cf.rules:
                # Check for a rule referencing column J (Rating) with >= 4
                if hasattr(rule, 'formula') and rule.formula:
                    for f in rule.formula:
                        if 'J' in f.upper() and ('4' in f or '>=4' in f):
                            rating_cf_found = True
                            break
                # Also check operator-based rules
                if hasattr(rule, 'operator') and rule.operator in ('greaterThanOrEqual', 'greaterThan'):
                    rating_cf_found = True

        if rating_cf_found:
            print(f"PASS: Component 4 — Conditional formatting for rating >= 4 found (0.10 pts)")
            total_score += 0.10
        else:
            # Check if there's any conditional formatting at all (partial)
            if len(cfs) > 0:
                print(f"PARTIAL: Component 4 — Conditional formatting exists but rating >=4 rule not confirmed (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 4 — No conditional formatting found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ---------------------------------------------------------------
    # Component 5: Summary sheet populated with formulas (0.20 points)
    # Task requires: books completed this year, average rating, books by genre (COUNTIF),
    #                ratings distribution
    # Initial Summary is empty; golden has COUNTIFS, AVERAGEIF, COUNTIF formulas
    # ---------------------------------------------------------------
    try:
        if 'Summary' not in wb.sheetnames:
            print(f"FAIL: Component 5 — 'Summary' sheet not found")
        else:
            ws_sum = wb['Summary']
            sub_score = 0.0

            # 5a: Check for COUNTIFS or COUNTIF formula for completed books (0.05)
            countifs_found = False
            averageif_found = False
            genre_countif_count = 0
            ratings_countif_count = 0

            for row in ws_sum.iter_rows(min_row=1, max_row=ws_sum.max_row, max_col=ws_sum.max_column):
                for cell in row:
                    if isinstance(cell.value, str):
                        val_upper = cell.value.upper()
                        if 'COUNTIFS' in val_upper or ('COUNTIF' in val_upper and 'COMPLETED' in val_upper.replace("'", "").replace('"', '')):
                            countifs_found = True
                        if 'AVERAGEIF' in val_upper:
                            averageif_found = True
                        if 'COUNTIF' in val_upper and 'D2' in cell.value.upper().replace("'READING LIST'!", ""):
                            genre_countif_count += 1
                        if 'COUNTIF' in val_upper and 'J2' in cell.value.upper().replace("'READING LIST'!", ""):
                            ratings_countif_count += 1

            # 5a: Completed count formula (0.05)
            if countifs_found:
                sub_score += 0.05
                print(f"  PASS: Component 5a — Books completed formula found")
            else:
                print(f"  FAIL: Component 5a — No completed books count formula")

            # 5b: Average rating formula (0.05)
            if averageif_found:
                sub_score += 0.05
                print(f"  PASS: Component 5b — Average rating formula found")
            else:
                print(f"  FAIL: Component 5b — No average rating formula")

            # 5c: Genre COUNTIF formulas (at least 5 of 7 genres) (0.05)
            if genre_countif_count >= 5:
                sub_score += 0.05
                print(f"  PASS: Component 5c — Genre COUNTIF formulas found ({genre_countif_count})")
            elif genre_countif_count > 0:
                partial = 0.05 * (genre_countif_count / 5)
                sub_score += partial
                print(f"  PARTIAL: Component 5c — Genre COUNTIF {genre_countif_count}/5+ ({partial:.3f} pts)")
            else:
                print(f"  FAIL: Component 5c — No genre COUNTIF formulas")

            # 5d: Ratings distribution COUNTIF (at least 3 of 5 ratings) (0.05)
            if ratings_countif_count >= 3:
                sub_score += 0.05
                print(f"  PASS: Component 5d — Ratings distribution COUNTIF found ({ratings_countif_count})")
            elif ratings_countif_count > 0:
                partial = 0.05 * (ratings_countif_count / 3)
                sub_score += partial
                print(f"  PARTIAL: Component 5d — Ratings COUNTIF {ratings_countif_count}/3+ ({partial:.3f} pts)")
            else:
                print(f"  FAIL: Component 5d — No ratings distribution COUNTIF formulas")

            print(f"SCORE: Component 5 — Summary formulas ({sub_score:.3f}/0.20 pts)")
            total_score += sub_score
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # ---------------------------------------------------------------
    # Component 6: Bar chart on Summary sheet (0.15 points)
    # Task requires: bar chart of ratings distribution
    # Initial Summary has 0 charts; golden has 1 bar chart
    # ---------------------------------------------------------------
    try:
        if 'Summary' in wb.sheetnames:
            ws_sum = wb['Summary']
            charts = ws_sum._charts
            if len(charts) >= 1:
                # Check if it's a bar/column chart
                chart = charts[0]
                chart_type = getattr(chart, 'type', '')
                if chart_type in ('col', 'bar'):
                    print(f"PASS: Component 6 — Bar chart found on Summary (type={chart_type}) (0.15 pts)")
                    total_score += 0.15
                else:
                    # Any chart is partial credit
                    print(f"PARTIAL: Component 6 — Chart found but type is '{chart_type}', expected bar/col (0.08 pts)")
                    total_score += 0.08
            else:
                print(f"FAIL: Component 6 — No charts found on Summary sheet")
        else:
            print(f"FAIL: Component 6 — Summary sheet not found")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # ---------------------------------------------------------------
    # Component 7: Sort order — Date Finished descending (0.10 points)
    # Task requires: Sort by Date Finished descending
    # Initial is sorted ascending; golden is sorted descending
    # We check that completed books (with dates) are in descending order
    # ---------------------------------------------------------------
    try:
        dates_finished = []
        for r in range(2, ws.max_row + 1):
            h_val = ws.cell(row=r, column=8).value  # H = Date Finished
            if h_val is not None:
                dates_finished.append(h_val)

        if len(dates_finished) >= 2:
            is_descending = all(dates_finished[i] >= dates_finished[i+1] for i in range(len(dates_finished)-1))
            if is_descending:
                print(f"PASS: Component 7 — Sorted by Date Finished descending ({len(dates_finished)} dates) (0.10 pts)")
                total_score += 0.10
            else:
                # Check if it's not the initial ascending order at least
                is_ascending = all(dates_finished[i] <= dates_finished[i+1] for i in range(len(dates_finished)-1))
                if is_ascending:
                    print(f"FAIL: Component 7 — Still in ascending order (initial state)")
                else:
                    print(f"FAIL: Component 7 — Not in descending order by Date Finished")
        else:
            print(f"FAIL: Component 7 — Not enough dates to check sort order ({len(dates_finished)} found)")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score:.4f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
