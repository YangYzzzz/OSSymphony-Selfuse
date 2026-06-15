"""
Initial Setup: Book Club Reading List Tracker
Task ID: calc_grs_038
Domain: libreoffice_calc

Creates a spreadsheet with book club reading data across two sheets.
The Reading List sheet has raw data but NO formulas, NO conditional formatting,
NO data validation dropdowns, and NO sorting. The Summary sheet exists but is empty.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_038'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Reading List ---
    ws = wb.active
    ws.title = "Reading List"

    # Headers
    headers = [
        "Book #", "Title", "Author", "Genre", "Year Published", "Pages",
        "Date Started", "Date Finished", "Days to Read", "Rating",
        "Maria's Review", "Jake's Review", "Priya's Review",
        "Tom's Review", "Lin's Review",
        "Average Member Rating", "Status"
    ]
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Book data: 15 books with realistic content
    # Columns: Book#, Title, Author, Genre, Year, Pages, DateStarted, DateFinished,
    #          DaysToRead(blank), Rating, Maria, Jake, Priya, Tom, Lin,
    #          AvgMemberRating(blank), Status
    books = [
        [1, "The Midnight Library", "Matt Haig", "Fiction", 2020, 288,
         date(2025, 1, 5), date(2025, 1, 18), None, 4,
         5, 4, 4, 3, 4, None, "Completed"],
        [2, "Atomic Habits", "James Clear", "Self-Help", 2018, 320,
         date(2025, 1, 22), date(2025, 2, 3), None, 5,
         5, 5, 4, 5, 5, None, "Completed"],
        [3, "Project Hail Mary", "Andy Weir", "Sci-Fi", 2021, 476,
         date(2025, 2, 10), date(2025, 2, 28), None, 5,
         5, 5, 5, 4, 5, None, "Completed"],
        [4, "The Silent Patient", "Alex Michaelides", "Mystery", 2019, 325,
         date(2025, 3, 1), date(2025, 3, 12), None, 3,
         3, 4, 2, 3, 3, None, "Completed"],
        [5, "Educated", "Tara Westover", "Biography", 2018, 334,
         date(2025, 3, 15), date(2025, 3, 29), None, 4,
         4, 5, 4, 4, 3, None, "Completed"],
        [6, "Dune", "Frank Herbert", "Sci-Fi", 1965, 412,
         date(2025, 4, 2), date(2025, 4, 20), None, 4,
         5, 4, 3, 5, 4, None, "Completed"],
        [7, "Where the Crawdads Sing", "Delia Owens", "Fiction", 2018, 368,
         date(2025, 4, 25), date(2025, 5, 8), None, 3,
         3, 2, 4, 3, 3, None, "Completed"],
        [8, "The Name of the Wind", "Patrick Rothfuss", "Fantasy", 2007, 662,
         date(2025, 5, 12), date(2025, 6, 5), None, 5,
         5, 5, 5, 5, 4, None, "Completed"],
        [9, "Sapiens", "Yuval Noah Harari", "Non-Fiction", 2011, 443,
         date(2025, 6, 10), date(2025, 6, 30), None, 4,
         4, 4, 5, 3, 4, None, "Completed"],
        [10, "The Hobbit", "J.R.R. Tolkien", "Fantasy", 1937, 310,
         date(2025, 7, 3), date(2025, 7, 15), None, 4,
         5, 4, 4, 4, 3, None, "Completed"],
        [11, "Becoming", "Michelle Obama", "Biography", 2018, 448,
         date(2025, 7, 20), date(2025, 8, 5), None, 3,
         3, 2, 4, 3, 3, None, "Completed"],
        [12, "The Alchemist", "Paulo Coelho", "Fiction", 1988, 197,
         date(2025, 8, 10), date(2025, 8, 18), None, 2,
         2, 3, 1, 2, 2, None, "Completed"],
        [13, "Neuromancer", "William Gibson", "Sci-Fi", 1984, 271,
         date(2025, 9, 1), None, None, None,
         None, None, None, None, None, None, "Currently Reading"],
        [14, "Gone Girl", "Gillian Flynn", "Mystery", 2012, 432,
         None, None, None, None,
         None, None, None, None, None, None, "To Read"],
        [15, "The Power of Now", "Eckhart Tolle", "Self-Help", 1997, 236,
         None, None, None, None,
         None, None, None, None, None, None, "To Read"],
    ]

    for r, row_data in enumerate(books, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c in (7, 8) and val is not None:
                cell.number_format = 'yyyy-mm-dd'
            if c == 1:
                cell.alignment = Alignment(horizontal="center")
            if c in (5, 6, 10, 11, 12, 13, 14, 15):
                cell.alignment = Alignment(horizontal="center")

    # Set column widths
    col_widths = {
        'A': 8, 'B': 30, 'C': 22, 'D': 14, 'E': 14, 'F': 8,
        'G': 14, 'H': 14, 'I': 13, 'J': 8,
        'K': 15, 'L': 14, 'M': 15, 'N': 14, 'O': 14,
        'P': 22, 'Q': 16
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # --- Sheet 2: Summary (empty, task asks agent to create content here) ---
    ws2 = wb.create_sheet("Summary")

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
