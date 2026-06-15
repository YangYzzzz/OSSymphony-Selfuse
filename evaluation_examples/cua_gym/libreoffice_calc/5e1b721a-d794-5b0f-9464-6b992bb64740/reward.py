"""
Reward Script: Lock formula cells and protect sheet with password
Task ID: calc_ps_041
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): Formula cells D2:D15 locked AND non-formula cells unlocked (selective locking)
  Component 2 (0.10): D17 grand total locked AND header D1 unlocked (selective locking of grand total)
  Component 3 (0.25): Sheet protection is enabled
  Component 4 (0.30): Sheet password matches 'sum2024'
"""

import os
import openpyxl
from openpyxl.worksheet.protection import SheetProtection

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_041'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Summary' sheet must exist
    if 'Summary' not in wb.sheetnames:
        print("FAIL: 'Summary' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Summary']

    # Component 1: Selective locking — D2:D15 locked AND sample non-formula cells unlocked (0.35 points)
    # This component verifies that formula cells STAY locked while other cells are UNLOCKED.
    # In initial_env, ALL cells are locked, so this combined check fails on initial (non-formula cells are locked).
    try:
        # Check formula cells D2:D15 are locked
        formula_locked_count = 0
        for row in range(2, 16):
            cell = ws.cell(row=row, column=4)
            if cell.protection.locked:
                formula_locked_count += 1

        # Check non-formula cells are unlocked
        non_formula_cells = ['A1', 'B1', 'C1', 'A2', 'B2', 'C2',
                             'A8', 'B8', 'C8', 'A15', 'B15', 'C15']
        unlocked_count = 0
        for coord in non_formula_cells:
            cell = ws[coord]
            if not cell.protection.locked:
                unlocked_count += 1

        all_formula_locked = (formula_locked_count == 14)
        all_nonformula_unlocked = (unlocked_count == len(non_formula_cells))

        if all_formula_locked and all_nonformula_unlocked:
            print(f"PASS: Component 1 — All 14 formula cells locked AND all {len(non_formula_cells)} "
                  f"non-formula cells unlocked (0.35 pts)")
            total_score += 0.35
        elif all_nonformula_unlocked and formula_locked_count >= 10:
            print(f"PARTIAL: Component 1 — {formula_locked_count}/14 formula cells locked, "
                  f"non-formula cells correctly unlocked (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 — formula locked: {formula_locked_count}/14, "
                  f"non-formula unlocked: {unlocked_count}/{len(non_formula_cells)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: D17 locked AND A17 unlocked (0.10 points)
    # Verifies selective locking for the grand total row.
    # In initial_env, both D17 and A17 are locked, so this fails.
    try:
        d17_locked = ws.cell(row=17, column=4).protection.locked
        a17_unlocked = not ws.cell(row=17, column=1).protection.locked

        if d17_locked and a17_unlocked:
            print(f"PASS: Component 2 — D17 locked and A17 unlocked (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 — D17 locked={d17_locked}, A17 unlocked={a17_unlocked}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Sheet protection is enabled (0.25 points)
    # In initial_env, sheet is NOT protected. This fails on initial.
    try:
        if ws.protection.sheet:
            print(f"PASS: Component 3 — Sheet protection is enabled (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 3 — Sheet is not protected")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Sheet password matches 'sum2024' (0.30 points)
    # In initial_env, no password is set. This fails on initial.
    try:
        prot = ws.protection
        if prot.sheet and prot.password:
            # Generate expected hash for 'sum2024'
            expected_sp = SheetProtection()
            expected_sp.set_password('sum2024')
            expected_hash = expected_sp.password

            if prot.password == expected_hash:
                print(f"PASS: Component 4 — Password hash matches 'sum2024' (0.30 pts)")
                total_score += 0.30
            else:
                print(f"FAIL: Component 4 — Password hash mismatch "
                      f"(expected {expected_hash}, got {prot.password})")
        else:
            print(f"FAIL: Component 4 — Sheet not protected or no password set")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
