"""
Initial Setup: Lock formula cells and protect Summary sheet
Task ID: calc_ps_041
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_041'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Summary'

    # --- Headers ---
    headers = ['Item', 'Qty', 'Price', 'Total']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # --- Data rows 2-15 (14 items) ---
    data = [
        ['Wireless Keyboard',       12,  45.99],
        ['USB-C Hub Adapter',       25,  32.50],
        ['LED Desk Lamp',            8,  67.00],
        ['Noise-Canceling Headset', 15,  89.95],
        ['Webcam HD 1080p',         20,  54.99],
        ['Ergonomic Mouse',         30,  29.99],
        ['Monitor Stand',           10,  42.00],
        ['HDMI Cable 6ft',          50,  12.99],
        ['Laptop Sleeve 15in',      18,  24.50],
        ['Wireless Charger',        22,  35.00],
        ['USB Flash Drive 64GB',    40,  15.99],
        ['Bluetooth Speaker',        6, 119.00],
        ['Screen Protector Pack',   35,   9.99],
        ['Power Strip Surge',       14,  27.50],
    ]
    for r, (item, qty, price) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=item)
        ws.cell(row=r, column=2, value=qty)
        ws.cell(row=r, column=3, value=price)
        # D column: formula =B*C
        ws.cell(row=r, column=4, value=f'=B{r}*C{r}')

    # --- Row 17: Grand Total ---
    ws.cell(row=17, column=1, value='Grand Total')
    ws.cell(row=17, column=4, value='=SUM(D2:D15)')

    # All cells are locked by default in openpyxl (Protection(locked=True))
    # Sheet is NOT protected - this is the default state

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
