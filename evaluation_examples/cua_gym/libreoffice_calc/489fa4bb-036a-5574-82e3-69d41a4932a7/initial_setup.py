"""
Initial Setup: Create pricing history spreadsheet for step chart task
Task ID: calc_gcp_056
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_056'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: PriceHistory ---
    ws = wb.active
    ws.title = 'PriceHistory'

    # Headers
    headers = ['EffectiveDate', 'Price']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal='center', vertical='center')
    white_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 15 dates spread irregularly across 2022-2024 with corresponding prices
    data = [
        (date(2022, 1, 15),  29.99),
        (date(2022, 3, 22),  29.99),
        (date(2022, 6, 10),  34.99),
        (date(2022, 8, 5),   34.99),
        (date(2022, 11, 18), 32.99),
        (date(2023, 2, 1),   39.99),
        (date(2023, 4, 14),  39.99),
        (date(2023, 7, 28),  44.99),
        (date(2023, 9, 12),  42.99),
        (date(2023, 11, 30), 42.99),
        (date(2024, 1, 20),  47.99),
        (date(2024, 3, 8),   47.99),
        (date(2024, 5, 25),  49.99),
        (date(2024, 8, 16),  52.99),
        (date(2024, 10, 3),  52.99),
    ]

    for r, (d, price) in enumerate(data, 2):
        date_cell = ws.cell(row=r, column=1, value=d)
        date_cell.number_format = 'yyyy-mm-dd'
        date_cell.alignment = Alignment(horizontal='center')

        price_cell = ws.cell(row=r, column=2, value=price)
        price_cell.number_format = '$#,##0.00'

    # Set column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 12

    # NO charts in initial state - that is the task for the agent

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
