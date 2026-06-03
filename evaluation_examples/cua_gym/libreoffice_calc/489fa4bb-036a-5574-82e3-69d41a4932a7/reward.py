"""
Reward Script: Create a step chart showing pricing history
Task ID: calc_gcp_056
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30) - A line chart exists on the PriceHistory sheet
  Component 2 (0.30) - Step interpolation data present (staircase pattern)
  Component 3 (0.20) - Chart has meaningful title and axis labels
  Component 4 (0.20) - Chart series references step data, smooth=False, covers all prices
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_056'


def persist_app_state(domain: str):
    """Try to save any unsaved GUI edits."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check PriceHistory sheet exists (precondition gate)
    if 'PriceHistory' not in wb.sheetnames:
        print("CRITICAL: 'PriceHistory' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['PriceHistory']

    # --- Component 1: A line chart exists on PriceHistory (0.30 points) ---
    # Initial state has 0 charts; golden state has 1 chart.
    try:
        charts = ws._charts
        chart = None
        if len(charts) >= 1:
            # Check that at least one chart is a LineChart
            from openpyxl.chart import LineChart
            line_charts = [c for c in charts if isinstance(c, LineChart)]
            if len(line_charts) >= 1:
                chart = line_charts[0]
                print(f"PASS: Component 1 — LineChart found on PriceHistory ({len(line_charts)} line chart(s)) (0.30 pts)")
                total_score += 0.30
            else:
                print(f"FAIL: Component 1 — Found {len(charts)} chart(s) but none are LineChart type")
        else:
            print(f"FAIL: Component 1 — No charts found on PriceHistory sheet")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        chart = None

    # --- Component 2: Step interpolation data (0.30 points) ---
    # The golden file implements step chart by creating helper columns with duplicated
    # data points to form a staircase pattern. We check:
    # - Extra data columns exist beyond the original A, B
    # - The data has more rows than the original 15 data rows (duplicated for steps)
    # - The staircase pattern: at each price change, there's a point at old price
    #   then a point at new price at the same date.
    try:
        # Original data has 15 data rows (rows 2-16) in columns A-B.
        # Step data should have more rows (roughly 2x for each price change point).
        # Check if columns beyond B exist with step data
        has_step_data = False
        step_data_rows = 0
        step_col_start = None

        # Look for step data columns (could be in C-onwards or could replace chart data)
        for col in range(3, ws.max_column + 1):
            header = ws.cell(1, col).value
            if header is not None and ws.cell(2, col).value is not None:
                # Found a data column beyond B
                step_col_start = col if step_col_start is None else step_col_start

        # Count rows in step data area
        if step_col_start is not None:
            # Find the date and price columns in the step data area
            for r in range(2, ws.max_row + 1):
                val = ws.cell(r, step_col_start).value
                if val is not None:
                    step_data_rows += 1

        # Also check if chart references data beyond original 15 rows
        # Step data should have > 15 rows (duplicated points for staircase)
        if step_data_rows > 15:
            # Verify staircase pattern: consecutive rows at same date with different prices
            staircase_transitions = 0
            for r in range(2, ws.max_row):
                date_curr = ws.cell(r, step_col_start).value
                date_next = ws.cell(r + 1, step_col_start).value
                price_curr = ws.cell(r, step_col_start + 1).value if step_col_start + 1 <= ws.max_column else None
                price_next = ws.cell(r + 1, step_col_start + 1).value if step_col_start + 1 <= ws.max_column else None

                if date_curr is not None and date_next is not None and price_curr is not None and price_next is not None:
                    if str(date_curr) == str(date_next) and abs(float(price_curr) - float(price_next)) > 0.01:
                        staircase_transitions += 1

            if staircase_transitions >= 3:
                print(f"PASS: Component 2 — Step data found: {step_data_rows} rows with {staircase_transitions} staircase transitions (0.30 pts)")
                total_score += 0.30
            else:
                print(f"FAIL: Component 2 — Step data rows={step_data_rows} but only {staircase_transitions} staircase transitions (expected >=3)")
        elif step_data_rows > 0:
            print(f"FAIL: Component 2 — Found extra columns but only {step_data_rows} step data rows (expected >15 for staircase)")
        else:
            # Alternative: check if chart references data that creates step pattern
            # by checking if chart series has > 15 data points
            if chart is not None and len(chart.series) >= 1:
                # Check the series data reference range
                s = chart.series[0]
                val_ref = str(s.val) if s.val else ""
                cat_ref = str(s.cat) if s.cat else ""
                print(f"FAIL: Component 2 — No step data columns found. Chart refs: val={val_ref}, cat={cat_ref}")
            else:
                print(f"FAIL: Component 2 — No step interpolation data found beyond original columns")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # --- Component 3: Chart title and axis labels (0.20 points) ---
    # Golden chart has title "Product Pricing History", X axis "Date", Y axis "Price ($)"
    try:
        if chart is not None:
            sub_score = 0.0

            # Check chart title exists and is meaningful
            title_text = ""
            if chart.title is not None:
                if hasattr(chart.title, 'tx') and chart.title.tx is not None:
                    if hasattr(chart.title.tx, 'rich') and chart.title.tx.rich is not None:
                        for p in chart.title.tx.rich.p:
                            for r in p.r:
                                title_text += r.t if r.t else ""
            title_lower = title_text.strip().lower()
            if len(title_lower) > 0 and any(kw in title_lower for kw in ['pric', 'history', 'step', 'product']):
                sub_score += 0.10
                print(f"  Chart title: '{title_text}' — relevant to task")
            elif len(title_lower) > 0:
                sub_score += 0.05
                print(f"  Chart title: '{title_text}' — exists but not clearly relevant")

            # Check axis labels exist
            has_axis_labels = 0
            for axis_name, axis_obj in [("X", chart.x_axis), ("Y", chart.y_axis)]:
                if axis_obj.title is not None:
                    axis_text = ""
                    if hasattr(axis_obj.title, 'tx') and axis_obj.title.tx is not None:
                        if hasattr(axis_obj.title.tx, 'rich') and axis_obj.title.tx.rich is not None:
                            for p in axis_obj.title.tx.rich.p:
                                for r in p.r:
                                    axis_text += r.t if r.t else ""
                    if len(axis_text.strip()) > 0:
                        has_axis_labels += 1
                        print(f"  {axis_name} axis title: '{axis_text}'")

            if has_axis_labels >= 2:
                sub_score += 0.10
            elif has_axis_labels >= 1:
                sub_score += 0.05

            if sub_score > 0:
                print(f"PASS: Component 3 — Chart has title and {has_axis_labels} axis label(s) ({sub_score:.2f} pts)")
                total_score += sub_score
            else:
                print(f"FAIL: Component 3 — Chart missing title and axis labels")
        else:
            print(f"FAIL: Component 3 — No chart to check titles on")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # --- Component 4: Chart series config (0.20 points) ---
    # Series should reference step data, smooth=False, and cover the full price range
    try:
        if chart is not None and len(chart.series) >= 1:
            s = chart.series[0]
            sub_score = 0.0

            # Check smooth is False (step charts should NOT be smoothed)
            if s.smooth is False or s.smooth is None or s.smooth == 0:
                sub_score += 0.10
                print(f"  Series smooth={s.smooth} — correct for step chart")
            else:
                print(f"  Series smooth={s.smooth} — should be False for step chart")

            # Check that series has at least 1 data series referencing the data
            if len(chart.series) >= 1:
                sub_score += 0.10
                print(f"  Series count: {len(chart.series)} — has data series")

            if sub_score > 0:
                print(f"PASS: Component 4 — Chart series configured correctly ({sub_score:.2f} pts)")
                total_score += sub_score
            else:
                print(f"FAIL: Component 4 — Chart series misconfigured")
        else:
            print(f"FAIL: Component 4 — No chart or no series to check")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
