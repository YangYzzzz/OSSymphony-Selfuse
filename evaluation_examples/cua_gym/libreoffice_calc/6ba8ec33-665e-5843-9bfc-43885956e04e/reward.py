"""
Reward Script: Set up a simple invoice template for freelance consulting business
Task ID: calc_gen_smallbiz_018
Domain: libreoffice_calc

Scoring Rubric (total 1.0):
  Component 1: Invoice header structure (business info + INVOICE label + invoice # + date formula) - 0.25 pts
  Component 2: Line-item table (row 6 headers + D7:D16 multiplication formulas) - 0.25 pts
  Component 3: Calculation section (Subtotal/Tax/Total labels + SUM + tax + total formulas) - 0.25 pts
  Component 4: Payment status dropdown (data validation on E21 with Unpaid/Paid/Overdue) - 0.25 pts
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path - all reward scripts run on the VM
TASK_ID = 'calc_gen_smallbiz_018'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: file must exist and be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Invoice sheet must exist
    if 'Invoice' not in wb.sheetnames:
        print("FAIL: No 'Invoice' sheet found in workbook")
        print(f"Score: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws = wb['Invoice']

    # -----------------------------------------------------------------------
    # Component 1: Invoice header structure (0.25 points)
    # Checks: B1='INVOICE', B2='Invoice #:', C2=1001 (starting number),
    #         B3='Date:', C3=TODAY() formula, plus business info in A1:A3
    # These are ALL absent in the initial file (empty sheet), so must be
    # task-introduced changes. Passes ONLY on golden, not on initial.
    # -----------------------------------------------------------------------
    try:
        b1_val = ws.cell(row=1, column=2).value
        b2_val = ws.cell(row=2, column=2).value
        c2_val = ws.cell(row=2, column=3).value
        b3_val = ws.cell(row=3, column=2).value
        c3_val = ws.cell(row=3, column=3).value
        a1_val = ws.cell(row=1, column=1).value

        b1_ok = isinstance(b1_val, str) and 'INVOICE' in b1_val.upper()
        b2_ok = isinstance(b2_val, str) and 'INVOICE' in b2_val.upper()
        c2_ok = c2_val is not None
        b3_ok = isinstance(b3_val, str) and 'DATE' in b3_val.upper()
        c3_ok = isinstance(c3_val, str) and 'TODAY()' in c3_val.upper()
        a1_ok = a1_val is not None and str(a1_val).strip() != ''

        header_checks = [b1_ok, b2_ok, c2_ok, b3_ok, c3_ok, a1_ok]
        header_passed = sum(header_checks)
        header_score = 0.25 if header_passed == len(header_checks) else (0.15 if header_passed >= 4 else 0.0)

        if header_score > 0.0:
            total_score += header_score
            print(f"PASS: Component 1 - Invoice header ({header_passed}/6 checks) ({header_score} pts): "
                  f"B1={repr(b1_val)}, B2={repr(b2_val)}, C2={repr(c2_val)}, C3={repr(c3_val)}, A1={repr(a1_val)}")
        else:
            print(f"FAIL: Component 1 - Invoice header missing ({header_passed}/6 checks): "
                  f"B1={repr(b1_val)}, B2={repr(b2_val)}, C2={repr(c2_val)}, "
                  f"B3={repr(b3_val)}, C3={repr(c3_val)}, A1={repr(a1_val)}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # -----------------------------------------------------------------------
    # Component 2: Line-item table structure (0.25 points)
    # Checks: row 6 headers (Description, Hours, Rate, Amount),
    #         D7:D16 all contain multiplication formulas (=Bn*Cn)
    # Passes ONLY on golden, not on initial.
    # -----------------------------------------------------------------------
    try:
        a6 = ws.cell(row=6, column=1).value
        b6 = ws.cell(row=6, column=2).value
        c6 = ws.cell(row=6, column=3).value
        d6 = ws.cell(row=6, column=4).value

        headers_ok = (
            isinstance(a6, str) and 'DESCRIPTION' in a6.upper() and
            isinstance(b6, str) and 'HOUR' in b6.upper() and
            isinstance(c6, str) and 'RATE' in c6.upper() and
            isinstance(d6, str) and 'AMOUNT' in d6.upper()
        )

        # Count D7:D16 cells that have the correct Bn*Cn multiplication formula
        formula_count = 0
        for r in range(7, 17):
            cell_val = ws.cell(row=r, column=4).value
            if isinstance(cell_val, str):
                v = cell_val.upper().replace(' ', '')
                if '*' in v and f'B{r}' in v and f'C{r}' in v:
                    formula_count += 1

        formulas_ok = formula_count == 10

        if headers_ok and formulas_ok:
            table_score = 0.25
        elif headers_ok or formula_count >= 8:
            table_score = (0.1 if headers_ok else 0.0) + (0.1 if formula_count >= 8 else 0.0)
        else:
            table_score = 0.0

        if table_score > 0.0:
            total_score += table_score
            print(f"PASS: Component 2 - Line-item table ({table_score} pts): "
                  f"headers_ok={headers_ok}, {formula_count}/10 multiplication formulas")
        else:
            print(f"FAIL: Component 2 - Line-item table: headers_ok={headers_ok}, "
                  f"{formula_count}/10 formulas, A6={repr(a6)}, B6={repr(b6)}, C6={repr(c6)}, D6={repr(d6)}")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # -----------------------------------------------------------------------
    # Component 3: Calculation section (0.25 points)
    # Checks: C17='Subtotal', D17=SUM(D7:D16), C18 has 8.5% tax label,
    #         D18=D17*0.085, C19='TOTAL', D19=D17+D18
    # Passes ONLY on golden, not on initial.
    # -----------------------------------------------------------------------
    try:
        c17 = ws.cell(row=17, column=3).value
        d17 = ws.cell(row=17, column=4).value
        c18 = ws.cell(row=18, column=3).value
        d18 = ws.cell(row=18, column=4).value
        c19 = ws.cell(row=19, column=3).value
        d19 = ws.cell(row=19, column=4).value

        c17_ok = isinstance(c17, str) and 'SUBTOTAL' in c17.upper()
        d17_ok = isinstance(d17, str) and 'SUM' in d17.upper() and 'D7' in d17.upper()
        c18_ok = isinstance(c18, str) and '8.5' in c18
        d18_ok = isinstance(d18, str) and '0.085' in d18 and 'D17' in d18.upper()
        c19_ok = isinstance(c19, str) and 'TOTAL' in c19.upper()
        d19_ok = isinstance(d19, str) and 'D17' in d19.upper() and 'D18' in d19.upper()

        calc_checks = [c17_ok, d17_ok, c18_ok, d18_ok, c19_ok, d19_ok]
        calc_passed = sum(calc_checks)
        calc_score = 0.25 if calc_passed == len(calc_checks) else (0.15 if calc_passed >= 4 else 0.0)

        if calc_score > 0.0:
            total_score += calc_score
            print(f"PASS: Component 3 - Calculation section ({calc_passed}/6, {calc_score} pts): "
                  f"C17={repr(c17)}, D17={repr(d17)}, C18={repr(c18)}, D18={repr(d18)}, "
                  f"C19={repr(c19)}, D19={repr(d19)}")
        else:
            print(f"FAIL: Component 3 - Calculation section ({calc_passed}/6 checks): "
                  f"C17={repr(c17)}, D17={repr(d17)}, C18={repr(c18)}, "
                  f"D18={repr(d18)}, C19={repr(c19)}, D19={repr(d19)}")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # -----------------------------------------------------------------------
    # Component 4: Payment status dropdown (0.25 points)
    # Checks: data validation of type 'list' on E21 with options
    #         Unpaid, Paid, Overdue
    # Passes ONLY on golden, not on initial (initial has no data validations).
    # -----------------------------------------------------------------------
    try:
        validations = ws.data_validations.dataValidation

        # Derived from real API inspection of each data validation object
        has_options = False
        has_cell_e21 = False
        found_formula = ''

        for dv in validations:
            if dv.type == 'list':
                formula = dv.formula1 if dv.formula1 else ''
                fu = formula.upper()
                options_ok = 'UNPAID' in fu and 'PAID' in fu and 'OVERDUE' in fu
                if options_ok:
                    has_options = options_ok
                    found_formula = formula
                    has_cell_e21 = 'E21' in str(dv.sqref).upper()

        if has_options and has_cell_e21:
            total_score += 0.25
            print(f"PASS: Component 4 - Payment dropdown (0.25 pts): "
                  f"formula={repr(found_formula)}, applied to E21")
        elif has_options:
            total_score += 0.15
            print(f"PARTIAL: Component 4 - Dropdown options correct but not on E21 (0.15 pts): "
                  f"formula={repr(found_formula)}")
        elif len(list(validations)) > 0:
            total_score += 0.10
            print("PARTIAL: Component 4 - Data validation exists but missing Unpaid/Paid/Overdue (0.10 pts)")
        else:
            print("FAIL: Component 4 - No list data validation found (payment dropdown missing)")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
