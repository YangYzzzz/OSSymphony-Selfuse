"""
Reward Script: Apply VLOOKUP approximate match for risk category lookup,
               sort by risk category, and add conditional formatting for critical risks.
Task ID: osworld_calc_vlookup_grade_lookup_007
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4 pts): Column C has VLOOKUP approximate match formulas referencing lookup table
  Component 2 (0.3 pts): Data rows are sorted by Risk Category (Critical → High → Low → Moderate)
  Component 3 (0.3 pts): Conditional formatting applied for 'Critical' rows with red background
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_vlookup_grade_lookup_007'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook (formula strings, not computed values)
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet 'Risk Data' must exist
    if 'Risk Data' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Risk Data' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Risk Data']

    # -------------------------------------------------------------------------
    # Component 1: Column C has VLOOKUP approximate match formulas (0.4 points)
    # Check that each data row (2-13) in column C has a VLOOKUP formula with
    # approximate match (last argument = 1 or TRUE), referencing the lookup table
    # in columns D:E.
    # -------------------------------------------------------------------------
    try:
        vlookup_count = 0
        approx_match_count = 0
        lookup_table_ref_count = 0
        total_data_rows = 12  # rows 2-13

        for row in range(2, total_data_rows + 2):
            cell_val = ws.cell(row=row, column=3).value
            if cell_val is None:
                continue
            formula_str = str(cell_val).strip()
            # Must be a formula
            if not formula_str.startswith('='):
                continue
            formula_upper = formula_str.upper()
            # Must contain VLOOKUP
            if 'VLOOKUP' not in formula_upper:
                continue
            vlookup_count += 1
            # Approximate match: last argument should be 1 or TRUE (not 0 or FALSE)
            # Pattern: ends with ,1) or ,TRUE) or ,1,) etc.
            if re.search(r',\s*(1|TRUE)\s*\)', formula_upper):
                approx_match_count += 1
            # Must reference the lookup table in D:E or $D$2:$E$5 range
            if re.search(r'\$?D\$?\d*:\$?E\$?\d*', formula_str, re.IGNORECASE) or \
               re.search(r'\$?D:\$?E', formula_str, re.IGNORECASE):
                lookup_table_ref_count += 1

        print(f"VLOOKUP formulas found in col C: {vlookup_count}/{total_data_rows}")
        print(f"  Approximate match (arg=1 or TRUE): {approx_match_count}")
        print(f"  References lookup table (D:E cols): {lookup_table_ref_count}")

        # Award 0.4 if all rows have VLOOKUP with approximate match and correct table ref
        if vlookup_count == total_data_rows and approx_match_count == total_data_rows and lookup_table_ref_count == total_data_rows:
            print("PASS: Component 1 — All 12 rows have VLOOKUP approximate match with correct table reference (0.4 pts)")
            total_score += 0.4
        elif vlookup_count >= total_data_rows and approx_match_count >= 1:
            # Partial: has VLOOKUPs but maybe not all approximate or not all rows
            partial = 0.2
            print(f"PARTIAL: Component 1 — VLOOKUP formulas present but incomplete criteria ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Expected VLOOKUP approximate match in all 12 col C cells, found {vlookup_count} VLOOKUPs, {approx_match_count} approximate")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Data is sorted by Risk Category (0.3 points)
    # The VLOOKUP assigns categories based on score thresholds:
    #   0-29 = Low, 30-59 = Moderate, 60-79 = High, 80+ = Critical
    # Expected sort order by category (alphabetical): Critical, High, Low, Moderate
    # We verify the risk scores in column B are grouped by their category bands
    # and the groups appear in sorted order (Critical first = scores >= 80,
    # then High = 60-79, then Low = 0-29, then Moderate = 30-59).
    # -------------------------------------------------------------------------
    try:
        # Extract the actual sorted project IDs and scores from the golden
        # Determine category from score using the thresholds in the task description:
        # 0=Low (0-29), 30=Moderate (30-59), 60=High (60-79), 80=Critical (80+)
        def score_to_category(score):
            if score is None:
                return None
            score = int(score)
            if score >= 80:
                return 'Critical'
            elif score >= 60:
                return 'High'
            elif score >= 30:
                return 'Moderate'
            else:
                return 'Low'

        data_rows = []
        for row in range(2, 14):
            score = ws.cell(row=row, column=2).value
            proj_id = ws.cell(row=row, column=1).value
            if score is not None and proj_id is not None:
                data_rows.append((proj_id, score, score_to_category(score)))

        if len(data_rows) < 12:
            print(f"FAIL: Component 2 — Only {len(data_rows)} data rows found, expected 12")
        else:
            # Extract categories in the current row order
            current_categories = [row[2] for row in data_rows]
            # Check that rows are sorted by category (consistent grouping)
            # Allowed sort order (alphabetical by category name): Critical < High < Low < Moderate
            sort_order = {'Critical': 0, 'High': 1, 'Low': 2, 'Moderate': 3}
            sorted_order_vals = [sort_order[c] for c in current_categories]
            is_sorted = all(sorted_order_vals[i] <= sorted_order_vals[i+1] for i in range(len(sorted_order_vals)-1))

            # Also verify that the sort is not just the original unsorted order
            # Original order was by Project ID (PRJ-2025-001 through PRJ-2025-012)
            original_ids = [f'PRJ-2025-{str(i).zfill(3)}' for i in range(1, 13)]
            current_ids = [row[0] for row in data_rows]
            is_original_order = (current_ids == original_ids)

            print(f"Category order in current file: {current_categories}")
            print(f"Is sorted by category (alphabetical): {is_sorted}")
            print(f"Is still in original Project ID order: {is_original_order}")

            if is_sorted and not is_original_order:
                print("PASS: Component 2 — Data is sorted by Risk Category (0.3 pts)")
                total_score += 0.3
            elif is_sorted and is_original_order:
                print("FAIL: Component 2 — Data appears sorted but is still in original order (unlikely coincidence)")
            else:
                print(f"FAIL: Component 2 — Data not sorted by Risk Category; found order: {current_categories}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Conditional formatting for 'Critical' rows with red background (0.3 points)
    # Expecting a formula-based conditional formatting rule on data range (e.g. A2:C13)
    # with formula like $C2="Critical" and fill color FFFF0000 (red).
    # -------------------------------------------------------------------------
    try:
        cf_rules_found = list(ws.conditional_formatting)
        print(f"\nConditional formatting ranges: {[str(r) for r in cf_rules_found]}")

        has_critical_cf = False
        has_red_fill = False

        for cf_range in cf_rules_found:
            for rule in ws.conditional_formatting[cf_range]:
                rule_type = getattr(rule, 'type', None)
                formula = getattr(rule, 'formula', None)
                dxf = getattr(rule, 'dxf', None)

                print(f"  CF rule: type={rule_type}, formula={formula}")

                # Check formula references 'Critical' in column C
                if formula:
                    formula_str = str(formula).upper()
                    if 'CRITICAL' in formula_str and ('C' in formula_str or '$C' in formula_str):
                        has_critical_cf = True
                        print(f"  Found: 'Critical' condition in col C formula: {formula}")

                # Check fill is red (FFFF0000)
                if dxf and dxf.fill:
                    try:
                        fill_color = dxf.fill.fgColor.rgb
                        print(f"  Fill color: {fill_color}")
                        if fill_color and fill_color.upper() in ('FFFF0000', 'FF0000', 'FF000000'):
                            has_red_fill = True
                            print(f"  Found: Red fill {fill_color}")
                    except Exception as color_err:
                        print(f"  Color check error: {color_err}")

        if has_critical_cf and has_red_fill:
            print("PASS: Component 3 — Conditional formatting for 'Critical' rows with red fill found (0.3 pts)")
            total_score += 0.3
        elif has_critical_cf:
            print("PARTIAL: Component 3 — Conditional formatting has 'Critical' formula but red fill not confirmed (0.15 pts)")
            total_score += 0.15
        elif has_red_fill:
            print("PARTIAL: Component 3 — Red fill CF rule found but 'Critical' condition not confirmed (0.15 pts)")
            total_score += 0.15
        else:
            print("FAIL: Component 3 — No conditional formatting for 'Critical' rows with red fill found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path on the VM
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
