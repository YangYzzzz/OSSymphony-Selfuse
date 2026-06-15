"""
Initial Setup: VLOOKUP approximate match to classify project risk scores into categories,
sort by risk category, and add conditional formatting to highlight critical risks.
Task ID: osworld_calc_vlookup_grade_lookup_007
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_vlookup_grade_lookup_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Risk Data ---
    ws = wb.active
    ws.title = 'Risk Data'

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18

    # Headers in A1:C1
    header_font = Font(bold=True)
    for col, header in enumerate(['Project ID', 'Risk Score', 'Risk Category'], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Risk lookup table headers in D1:E1
    for col, header in enumerate(['Score Threshold', 'Category'], 4):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Risk lookup table data in D2:E5 (must be sorted ascending for VLOOKUP approximate match)
    lookup_data = [
        (0, 'Low'),
        (30, 'Moderate'),
        (60, 'High'),
        (80, 'Critical'),
    ]
    for r, (threshold, category) in enumerate(lookup_data, 2):
        ws.cell(row=r, column=4, value=threshold)
        ws.cell(row=r, column=5, value=category)

    # Project data in A2:B13 — intentionally NOT sorted by category (mixed order)
    # Risk scores are varied so VLOOKUP will map to all four categories
    # Column C (Risk Category) is intentionally EMPTY — that's the task
    project_data = [
        ('PRJ-2025-001', 85),   # Critical
        ('PRJ-2025-002', 42),   # Moderate
        ('PRJ-2025-003', 15),   # Low
        ('PRJ-2025-004', 73),   # High
        ('PRJ-2025-005', 91),   # Critical
        ('PRJ-2025-006', 55),   # Moderate
        ('PRJ-2025-007', 8),    # Low
        ('PRJ-2025-008', 68),   # High
        ('PRJ-2025-009', 35),   # Moderate
        ('PRJ-2025-010', 82),   # Critical
        ('PRJ-2025-011', 25),   # Low
        ('PRJ-2025-012', 77),   # High
    ]
    for r, (project_id, risk_score) in enumerate(project_data, 2):
        ws.cell(row=r, column=1, value=project_id)
        ws.cell(row=r, column=2, value=risk_score)
        # Column C (Risk Category) is left EMPTY — task is to add VLOOKUP here

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
