"""
Initial Setup: Remove sheet protection from Budget sheet
Task ID: calc_gsi_017
Domain: libreoffice_calc

Creates a Q2 budget workbook with the 'Budget' sheet protected by password 'secure123'.
The agent must unprotect it and update the travel expense in D15.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_017'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Budget (Q2 Budget) ---
    ws_budget = wb.active
    ws_budget.title = 'Budget'

    # Header styling
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    currency_fmt = '$#,##0.00'
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    # Column headers
    headers = ['Category', 'Q1 Actual', 'Q2 Budget', 'Q2 Actual', 'Variance']
    for col, h in enumerate(headers, 1):
        cell = ws_budget.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Budget line items: (Category, Q1 Actual, Q2 Budget, Q2 Actual, Variance formula)
    budget_data = [
        ('Salaries & Wages', 245000.00, 252000.00, 251800.00),
        ('Employee Benefits', 73500.00, 75600.00, 74950.00),
        ('Contractor Fees', 35000.00, 42000.00, 39500.00),
        ('Office Rent', 18000.00, 18000.00, 18000.00),
        ('Utilities', 4200.00, 4500.00, 4380.00),
        ('Office Supplies', 2800.00, 3200.00, 2950.00),
        ('Software Licenses', 15600.00, 16800.00, 16200.00),
        ('Hardware & Equipment', 8500.00, 12000.00, 9800.00),
        ('Marketing - Digital', 22000.00, 28000.00, 26500.00),
        ('Marketing - Events', 15000.00, 18000.00, 17200.00),
        ('Professional Development', 6000.00, 8000.00, 7500.00),
        ('Insurance', 9200.00, 9200.00, 9200.00),
        ('Telecommunications', 3600.00, 3800.00, 3750.00),
        ('Travel Expenses', 14000.00, 16000.00, 12500.00),   # Row 15 (data row 14, row offset 2+13=15)
        ('Client Entertainment', 5500.00, 7000.00, 6200.00),
        ('Legal & Compliance', 8000.00, 8500.00, 8100.00),
        ('Printing & Postage', 1800.00, 2000.00, 1750.00),
        ('Maintenance & Repairs', 4500.00, 5000.00, 4600.00),
        ('Miscellaneous', 3000.00, 3500.00, 3200.00),
    ]

    for r, (cat, q1, q2b, q2a) in enumerate(budget_data, 2):
        ws_budget.cell(row=r, column=1, value=cat).border = thin_border
        ws_budget.cell(row=r, column=2, value=q1).number_format = currency_fmt
        ws_budget.cell(row=r, column=2).border = thin_border
        ws_budget.cell(row=r, column=3, value=q2b).number_format = currency_fmt
        ws_budget.cell(row=r, column=3).border = thin_border
        ws_budget.cell(row=r, column=4, value=q2a).number_format = currency_fmt
        ws_budget.cell(row=r, column=4).border = thin_border
        # Variance = Q2 Actual - Q2 Budget
        var_cell = ws_budget.cell(row=r, column=5, value=f'=D{r}-C{r}')
        var_cell.number_format = currency_fmt
        var_cell.border = thin_border

    # Totals row
    total_row = len(budget_data) + 2  # row 21
    ws_budget.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    ws_budget.cell(row=total_row, column=1).border = thin_border
    for col in range(2, 6):
        col_letter = openpyxl.utils.get_column_letter(col)
        total_cell = ws_budget.cell(row=total_row, column=col,
                                     value=f'=SUM({col_letter}2:{col_letter}{total_row-1})')
        total_cell.number_format = currency_fmt
        total_cell.font = Font(bold=True)
        total_cell.border = thin_border

    # Column widths
    ws_budget.column_dimensions['A'].width = 28
    ws_budget.column_dimensions['B'].width = 16
    ws_budget.column_dimensions['C'].width = 16
    ws_budget.column_dimensions['D'].width = 16
    ws_budget.column_dimensions['E'].width = 16

    # Freeze header row
    ws_budget.freeze_panes = 'A2'

    # PROTECT the Budget sheet with password 'secure123'
    ws_budget.protection.sheet = True
    ws_budget.protection.password = 'secure123'

    # --- Sheet 2: Summary ---
    ws_summary = wb.create_sheet('Summary')

    ws_summary['A1'] = 'Fiscal Year 2025-2026 Budget Overview'
    ws_summary['A1'].font = Font(name='Arial', size=14, bold=True)

    ws_summary['A3'] = 'Department:'
    ws_summary['B3'] = 'Finance & Operations'
    ws_summary['A4'] = 'Prepared by:'
    ws_summary['B4'] = 'Rachel Martinez'
    ws_summary['A5'] = 'Last Updated:'
    ws_summary['B5'] = '2025-09-28'
    ws_summary['A6'] = 'Status:'
    ws_summary['B6'] = 'Under Review'

    summary_labels = ['A3', 'A4', 'A5', 'A6']
    for coord in summary_labels:
        ws_summary[coord].font = Font(bold=True)

    ws_summary['A8'] = 'Notes:'
    ws_summary['A8'].font = Font(bold=True, italic=True)
    ws_summary['A9'] = 'The Budget sheet is currently protected to prevent accidental edits.'
    ws_summary['A10'] = 'Contact the finance team for the unlock password if updates are needed.'
    ws_summary['A11'] = 'Q2 travel expenses need to be reconciled with submitted receipts.'

    ws_summary.column_dimensions['A'].width = 22
    ws_summary.column_dimensions['B'].width = 30

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
