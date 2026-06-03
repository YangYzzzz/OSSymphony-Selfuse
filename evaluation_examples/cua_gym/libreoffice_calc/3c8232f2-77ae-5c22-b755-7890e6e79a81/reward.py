"""
Reward Script: Multi-app data workflow with survey.ods
Task ID: osworld_multi_apps_multi_case_convert_008
Domain: libreoffice_calc (multi-app workflow)
Scoring:
  Component 1 (0.30): clean.csv exists, 40 rows, no NaN, score cols normalized [0,1]
  Component 2 (0.25): survey.xlsx exists with exactly 2 sheets
  Component 3 (0.20): report.html exists and contains an HTML <table>
  Component 4 (0.10): survey.pdf exists with non-zero size
  Component 5 (0.15): workflow.log exists and contains step entries with file sizes
  Total: 1.0
"""

import os

DATA_DIR = '/home/user/data'
TASK_ID = 'osworld_multi_apps_multi_case_convert_008'

SCORE_COLS = [
    'satisfaction_score', 'engagement_score', 'productivity_score',
    'collaboration_score', 'innovation_score', 'wellbeing_score'
]

def verify_task():
    """
    Verify multi-app workflow task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # -----------------------------------------------------------------------
    # Component 1: clean.csv — exists, 40 rows, no NaN, score cols in [0,1]
    # (0.30 points)
    # -----------------------------------------------------------------------
    try:
        import pandas as pd

        clean_csv_path = os.path.join(DATA_DIR, 'clean.csv')
        if not os.path.exists(clean_csv_path):
            print(f"FAIL: Component 1 — clean.csv not found at {clean_csv_path}")
        else:
            df = pd.read_csv(clean_csv_path)

            # Row count check
            row_ok = len(df) == 40
            if not row_ok:
                print(f"FAIL: Component 1 — clean.csv has {len(df)} rows, expected 40")

            # No NaN check
            nan_total = df.isnull().sum().sum()
            nan_ok = nan_total == 0
            if not nan_ok:
                print(f"FAIL: Component 1 — clean.csv has {nan_total} NaN values, expected 0")

            # Score columns normalized to [0, 1]
            score_cols_present = [c for c in SCORE_COLS if c in df.columns]
            norm_ok = len(score_cols_present) == len(SCORE_COLS)
            if norm_ok:
                for col in score_cols_present:
                    col_min = df[col].min()
                    col_max = df[col].max()
                    if col_min < -0.001 or col_max > 1.001:
                        norm_ok = False
                        print(f"FAIL: Component 1 — {col} not in [0,1]: min={col_min:.4f}, max={col_max:.4f}")
                        break
            else:
                print(f"FAIL: Component 1 — score columns missing; found {score_cols_present}")

            if row_ok and nan_ok and norm_ok:
                print(f"PASS: Component 1 — clean.csv: 40 rows, no NaN, all {len(SCORE_COLS)} score cols normalized [0,1] (0.30 pts)")
                total_score += 0.30
            else:
                print(f"FAIL: Component 1 — clean.csv quality checks failed (row_ok={row_ok}, nan_ok={nan_ok}, norm_ok={norm_ok})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: survey.xlsx — exists, has exactly 2 sheets
    # (0.25 points)
    # -----------------------------------------------------------------------
    try:
        import openpyxl

        xlsx_path = os.path.join(DATA_DIR, 'survey.xlsx')
        if not os.path.exists(xlsx_path):
            print(f"FAIL: Component 2 — survey.xlsx not found at {xlsx_path}")
        else:
            wb = openpyxl.load_workbook(xlsx_path)
            sheet_count = len(wb.sheetnames)
            if sheet_count == 2:
                print(f"PASS: Component 2 — survey.xlsx has exactly 2 sheets: {wb.sheetnames} (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 2 — survey.xlsx has {sheet_count} sheet(s): {wb.sheetnames}, expected 2")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: report.html — exists and contains an HTML <table> element
    # (0.20 points)
    # -----------------------------------------------------------------------
    try:
        html_path = os.path.join(DATA_DIR, 'report.html')
        if not os.path.exists(html_path):
            print(f"FAIL: Component 3 — report.html not found at {html_path}")
        else:
            with open(html_path, 'r', encoding='utf-8', errors='replace') as f:
                html_content = f.read()

            has_table = '<table' in html_content.lower()
            has_html_tag = '<html' in html_content.lower() or '<!doctype html' in html_content.lower()

            if has_table and has_html_tag:
                print(f"PASS: Component 3 — report.html is a valid HTML file with a <table> element (0.20 pts)")
                total_score += 0.20
            elif has_table:
                # Partial: has table but not standard HTML structure
                print(f"PASS (partial): Component 3 — report.html has <table> but no <html> tag; awarding partial (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 3 — report.html exists but has no <table> element (has_table={has_table}, has_html={has_html_tag})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: survey.pdf — exists with non-zero size
    # (0.10 points)
    # -----------------------------------------------------------------------
    try:
        pdf_path = os.path.join(DATA_DIR, 'survey.pdf')
        if not os.path.exists(pdf_path):
            print(f"FAIL: Component 4 — survey.pdf not found at {pdf_path}")
        else:
            pdf_size = os.path.getsize(pdf_path)
            if pdf_size > 0:
                # Verify it at least starts with %PDF header
                with open(pdf_path, 'rb') as f:
                    header = f.read(4)
                if header == b'%PDF':
                    print(f"PASS: Component 4 — survey.pdf exists ({pdf_size} bytes), valid PDF header (0.10 pts)")
                    total_score += 0.10
                else:
                    print(f"FAIL: Component 4 — survey.pdf exists ({pdf_size} bytes) but does not have a PDF header (got {header})")
            else:
                print(f"FAIL: Component 4 — survey.pdf exists but is empty (0 bytes)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -----------------------------------------------------------------------
    # Component 5: workflow.log — exists and contains step entries with file sizes
    # (0.15 points)
    # -----------------------------------------------------------------------
    try:
        log_path = os.path.join(DATA_DIR, 'workflow.log')
        if not os.path.exists(log_path):
            print(f"FAIL: Component 5 — workflow.log not found at {log_path}")
        else:
            with open(log_path, 'r', encoding='utf-8', errors='replace') as f:
                log_content = f.read()

            # Check: log has multiple lines (steps)
            log_lines = [l for l in log_content.strip().split('\n') if l.strip()]
            has_multiple_steps = len(log_lines) >= 3
            # Check: log records file sizes (bytes keyword or size: xxx pattern)
            has_file_sizes = 'bytes' in log_content.lower() or 'size' in log_content.lower()
            # Check: log references key files
            references_files = any(f in log_content for f in ['clean.csv', 'survey.xlsx', 'report.html', 'survey.pdf'])

            if has_multiple_steps and has_file_sizes and references_files:
                print(f"PASS: Component 5 — workflow.log has {len(log_lines)} step entries, records file sizes, references output files (0.15 pts)")
                total_score += 0.15
            elif has_multiple_steps and has_file_sizes:
                print(f"PASS (partial): Component 5 — workflow.log has steps and sizes but may lack file references; awarding 0.15 pts")
                total_score += 0.15
            else:
                print(f"FAIL: Component 5 — workflow.log incomplete (steps={has_multiple_steps}, sizes={has_file_sizes}, refs={references_files})")
                print(f"  Log content preview: {log_content[:200]}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


verify_task()
