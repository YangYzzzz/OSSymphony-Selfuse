"""
Initial Setup: Column chart with 9pt axis tick label font sizes
Task ID: calc_chart_column_font_size_038
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.chart import BarChart, Reference
from lxml import etree

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_column_font_size_038'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def set_axis_txPr(axis_obj, sz):
    """Set the tick label font size on a chart axis object via txPr XML.
    sz: font size * 100 (e.g., 900 = 9pt, 1200 = 12pt)
    """
    a_ns = 'http://schemas.openxmlformats.org/drawingml/2006/main'

    # Build a txPr element with defRPr sz attribute
    txPr_xml = (
        f'<txPr xmlns:a="{a_ns}">'
        f'<a:bodyPr/>'
        f'<a:lstStyle/>'
        f'<a:p><a:pPr><a:defRPr sz="{sz}" b="0"/></a:pPr></a:p>'
        f'</txPr>'
    )
    txPr_elem = etree.fromstring(txPr_xml)

    # Assign it to the axis's txPr attribute (openpyxl Axis uses _txPr or txPr)
    # openpyxl wraps axis properties; we set via the object's element namespace
    axis_obj.txPr = txPr_elem


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'PrintReport'

    # Headers
    ws['A1'] = 'Product Line'
    ws['B1'] = 'Annual Sales'

    # Data rows
    data = [
        ('Electronics', 2840000),
        ('Appliances',  1920000),
        ('Furniture',   1380000),
        ('Clothing',     890000),
        ('Sports',       720000),
    ]
    for r, (product, sales) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=product)
        ws.cell(row=r, column=2, value=sales)

    # Create a column (vertical bar) chart
    chart = BarChart()
    chart.type = 'col'
    chart.title = 'Annual Sales by Product Line'
    chart.y_axis.title = 'Annual Sales ($)'
    chart.x_axis.title = 'Product Line'

    data_ref = Reference(ws, min_col=2, min_row=1, max_row=6)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=6)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    chart.shape = 4
    chart.width = 18
    chart.height = 12

    ws.add_chart(chart, 'D2')

    # Save first so we can reload and manipulate the XML
    wb.save(OUTPUT)

    # Reload and set axis tick label font sizes to 9pt via XML manipulation
    import zipfile
    import io

    with zipfile.ZipFile(OUTPUT, 'r') as zin:
        names = zin.namelist()
        files = {}
        for name in names:
            files[name] = zin.read(name)

    # Find chart XML files
    chart_files = [n for n in names if n.startswith('xl/charts/chart')]
    print(f'Chart files found: {chart_files}')

    c_ns = 'http://schemas.openxmlformats.org/drawingml/2006/chart'
    a_ns = 'http://schemas.openxmlformats.org/drawingml/2006/main'

    def set_axis_txPr_in_xml(axis_elem, sz):
        """Set txPr on an axis element with given font size (sz = size*100)."""
        # Remove existing txPr if any
        existing = axis_elem.find(f'{{{c_ns}}}txPr')
        if existing is not None:
            axis_elem.remove(existing)

        # Build new txPr
        txPr = etree.SubElement(axis_elem, f'{{{c_ns}}}txPr')
        bodyPr = etree.SubElement(txPr, f'{{{a_ns}}}bodyPr')
        lstStyle = etree.SubElement(txPr, f'{{{a_ns}}}lstStyle')
        p = etree.SubElement(txPr, f'{{{a_ns}}}p')
        pPr = etree.SubElement(p, f'{{{a_ns}}}pPr')
        defRPr = etree.SubElement(pPr, f'{{{a_ns}}}defRPr')
        defRPr.set('sz', str(sz))
        defRPr.set('b', '0')

    for chart_file in chart_files:
        tree = etree.fromstring(files[chart_file])
        plotArea = tree.find(
            f'{{{c_ns}}}chart/{{{c_ns}}}plotArea'
        )
        if plotArea is not None:
            for catAx in plotArea.findall(f'{{{c_ns}}}catAx'):
                set_axis_txPr_in_xml(catAx, 900)   # 9pt
            for valAx in plotArea.findall(f'{{{c_ns}}}valAx'):
                set_axis_txPr_in_xml(valAx, 900)   # 9pt

        files[chart_file] = etree.tostring(tree, xml_declaration=True,
                                           encoding='UTF-8', standalone=True)

    # Write modified zip
    with zipfile.ZipFile(OUTPUT, 'w', zipfile.ZIP_DEFLATED) as zout:
        for name, data in files.items():
            zout.writestr(name, data)

    print(f'Initial file created: {OUTPUT}')
    print('Chart created with 9pt axis tick label font sizes (both X and Y axes).')


create_initial()
