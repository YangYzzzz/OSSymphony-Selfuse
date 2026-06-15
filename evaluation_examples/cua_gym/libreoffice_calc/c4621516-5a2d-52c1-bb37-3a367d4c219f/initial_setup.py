"""
Initial Setup: Inventory Turnover Table with 6 Product Lines and 3 Years of Data
Task ID: osworld_calc_annual_pct_change_008
Domain: libreoffice_calc

Layout: Row 1 = headers (Metric/Product | Product columns B-G)
        Row 2 = 2021 turnover rates
        Row 3 = 2022 turnover rates
        Row 4 = 2023 turnover rates
        (No change rows or ranking — those are added by the task)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_annual_pct_change_008'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventory Turnover'

    # Style definitions
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_font_white = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    data_font = Font(name='Calibri', size=11)
    data_bold_font = Font(name='Calibri', bold=True, size=11)
    data_align_left = Alignment(horizontal='left', vertical='center')
    data_align_center = Alignment(horizontal='center', vertical='center')

    # Product line names (as column headers B-G)
    products = [
        'Electronics',
        'Apparel',
        'Home & Garden',
        'Sports Equipment',
        'Automotive Parts',
        'Office Supplies',
    ]

    # Turnover data: [product_idx][year_idx]  year: 0=2021, 1=2022, 2=2023
    turnover = [
        [8.4,  9.1,  10.2],   # Electronics
        [5.2,  4.8,   5.5],   # Apparel
        [4.1,  4.6,   5.0],   # Home & Garden
        [6.3,  7.0,   6.8],   # Sports Equipment
        [3.8,  4.2,   4.9],   # Automotive Parts
        [7.5,  8.3,   9.1],   # Office Supplies
    ]

    # Row 1: Headers — col A = "Metric / Product", cols B-G = product names
    ws.cell(row=1, column=1, value='Year / Product').font = header_font_white
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = header_align
    ws.cell(row=1, column=1).border = border

    for col, pname in enumerate(products, 2):
        cell = ws.cell(row=1, column=col, value=pname)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Row 2: 2021 values
    ws.cell(row=2, column=1, value='2021').font = data_bold_font
    ws.cell(row=2, column=1).alignment = data_align_left
    ws.cell(row=2, column=1).border = border
    for col, i in enumerate(range(6), 0):
        cell = ws.cell(row=2, column=col+2, value=turnover[i][0])
        cell.font = data_font
        cell.alignment = data_align_center
        cell.number_format = '0.0'
        cell.border = border

    # Row 3: 2022 values
    ws.cell(row=3, column=1, value='2022').font = data_bold_font
    ws.cell(row=3, column=1).alignment = data_align_left
    ws.cell(row=3, column=1).border = border
    for col, i in enumerate(range(6), 0):
        cell = ws.cell(row=3, column=col+2, value=turnover[i][1])
        cell.font = data_font
        cell.alignment = data_align_center
        cell.number_format = '0.0'
        cell.border = border

    # Row 4: 2023 values
    ws.cell(row=4, column=1, value='2023').font = data_bold_font
    ws.cell(row=4, column=1).alignment = data_align_left
    ws.cell(row=4, column=1).border = border
    for col, i in enumerate(range(6), 0):
        cell = ws.cell(row=4, column=col+2, value=turnover[i][2])
        cell.font = data_font
        cell.alignment = data_align_center
        cell.number_format = '0.0'
        cell.border = border

    # Column widths
    ws.column_dimensions['A'].width = 22
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].width = 18

    # Row heights
    for r in range(1, 5):
        ws.row_dimensions[r].height = 20

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the spreadsheet in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
