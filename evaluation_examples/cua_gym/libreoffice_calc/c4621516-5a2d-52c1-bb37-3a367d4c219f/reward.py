"""
Reward Script: Inventory Turnover Analysis — Absolute Change, % Change, and Ranking
Task ID: osworld_calc_annual_pct_change_008
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Absolute change rows (2021→2022 and 2022→2023) present and correct — 0.30 pts
  Component 2: Percentage change rows (2021→2022 and 2022→2023) present and correct — 0.40 pts
  Component 3: Ranking row present and correct (by avg YoY growth rate) — 0.30 pts
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_annual_pct_change_008'

# Expected values (ground truth derived from task requirements and initial data)
# Initial data: rows 2-4, cols B-G (6 product lines, 3 years)
# Products (cols B-G): Electronics, Apparel, Home & Garden, Sports Equipment, Automotive Parts, Office Supplies
# Year 2021 (row 2): 8.4, 5.2, 4.1, 6.3, 3.8, 7.5
# Year 2022 (row 3): 9.1, 4.8, 4.6, 7.0, 4.2, 8.3
# Year 2023 (row 4): 10.2, 5.5, 5.0, 6.8, 4.9, 9.1

EXPECTED_ABS_2021_2022 = [0.7, -0.4, 0.5, 0.7, 0.4, 0.8]   # row2022 - row2021, cols B-G
EXPECTED_ABS_2022_2023 = [1.1, 0.7, 0.4, -0.2, 0.7, 0.8]   # row2023 - row2022, cols B-G

# % change = abs_change / base_year, stored as decimal (0.0833 = 8.33%)
EXPECTED_PCT_2021_2022 = [
    round(0.7 / 8.4, 4),    # Electronics
    round(-0.4 / 5.2, 4),   # Apparel
    round(0.5 / 4.1, 4),    # Home & Garden
    round(0.7 / 6.3, 4),    # Sports Equipment
    round(0.4 / 3.8, 4),    # Automotive Parts
    round(0.8 / 7.5, 4),    # Office Supplies
]
EXPECTED_PCT_2022_2023 = [
    round(1.1 / 9.1, 4),    # Electronics
    round(0.7 / 4.8, 4),    # Apparel
    round(0.4 / 4.6, 4),    # Home & Garden
    round(-0.2 / 7.0, 4),   # Sports Equipment
    round(0.7 / 4.2, 4),    # Automotive Parts
    round(0.8 / 8.3, 4),    # Office Supplies
]

# Rankings by avg YoY growth (rank 1 = highest avg growth)
# Avg growth rates: Automotive Parts(0.1360), Home&Garden(0.1045), Electronics(0.1021),
#                  Office Supplies(0.1015), Sports Equipment(0.0413), Apparel(0.0345)
EXPECTED_RANKS = [3, 6, 2, 5, 1, 4]  # cols B-G


def close_enough(a, b, tol=0.005):
    """Check if two numeric values are within tolerance."""
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def find_row_by_label(ws, label_substr):
    """Find the row number whose column A contains the given label substring (case-insensitive)."""
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if cell.value and label_substr.lower() in str(cell.value).lower():
            return cell.row
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook — precondition gate
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get active sheet
    try:
        ws = wb.active
        print(f"INFO: Active sheet = '{ws.title}', max_row={ws.max_row}, max_col={ws.max_column}")
    except Exception as e:
        print(f"CRITICAL: Cannot access active sheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: verify original data rows (rows 1-4) are intact
    try:
        header_ok = ws['A1'].value is not None
        data_rows_present = ws.max_row >= 4
        if not header_ok or not data_rows_present:
            print("CRITICAL: Original data rows (1-4) appear missing or corrupted")
            print("REWARD: 0.0")
            return 0.0
    except Exception as e:
        print(f"CRITICAL: Cannot verify original data: {e}")
        print("REWARD: 0.0")
        return 0.0

    # -------------------------------------------------------------------
    # Component 1: Absolute change rows present and correct (0.30 points)
    # Abs Change 2021→2022 row and Abs Change 2022→2023 row must exist
    # with values matching year2 - year1 for each product column
    # -------------------------------------------------------------------
    try:
        abs_row1 = find_row_by_label(ws, '2021')
        abs_row2 = find_row_by_label(ws, '2022')

        # Look for absolute change rows
        abs_change_row1 = None
        abs_change_row2 = None
        for row in ws.iter_rows(min_col=1, max_col=1):
            cell = row[0]
            if cell.value:
                label = str(cell.value).lower()
                if 'abs' in label and '2021' in label and '2022' in label:
                    abs_change_row1 = cell.row
                elif 'abs' in label and '2022' in label and '2023' in label:
                    abs_change_row2 = cell.row

        if abs_change_row1 is None or abs_change_row2 is None:
            print(f"FAIL: Component 1 — Absolute change rows not found (abs_row1={abs_change_row1}, abs_row2={abs_change_row2})")
        else:
            # Verify values for Abs Change 2021→2022 (cols B-G = cols 2-7)
            abs1_correct = 0
            abs1_total = 6
            for col_idx in range(2, 8):  # B to G
                actual = ws.cell(row=abs_change_row1, column=col_idx).value
                expected = EXPECTED_ABS_2021_2022[col_idx - 2]
                if close_enough(actual, expected, tol=0.02):
                    abs1_correct += 1
                else:
                    print(f"  FAIL abs 2021->2022 col {col_idx}: expected {expected}, got {actual}")

            # Verify values for Abs Change 2022→2023 (cols B-G = cols 2-7)
            abs2_correct = 0
            abs2_total = 6
            for col_idx in range(2, 8):  # B to G
                actual = ws.cell(row=abs_change_row2, column=col_idx).value
                expected = EXPECTED_ABS_2022_2023[col_idx - 2]
                if close_enough(actual, expected, tol=0.02):
                    abs2_correct += 1
                else:
                    print(f"  FAIL abs 2022->2023 col {col_idx}: expected {expected}, got {actual}")

            total_correct = abs1_correct + abs2_correct
            total_possible = abs1_total + abs2_total

            if total_correct == total_possible:
                print(f"PASS: Component 1 — Both absolute change rows correct ({total_correct}/{total_possible} values) (0.30 pts)")
                total_score += 0.30
            elif total_correct >= total_possible * 0.5:
                print(f"PARTIAL: Component 1 — {total_correct}/{total_possible} absolute change values correct (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 1 — Only {total_correct}/{total_possible} absolute change values correct")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------
    # Component 2: Percentage change rows present and correct (0.40 points)
    # % Change 2021→2022 and % Change 2022→2023 rows must exist
    # with values as decimal fractions (e.g., 0.0833 = 8.33%)
    # -------------------------------------------------------------------
    try:
        pct_change_row1 = None
        pct_change_row2 = None
        for row in ws.iter_rows(min_col=1, max_col=1):
            cell = row[0]
            if cell.value:
                label = str(cell.value).lower()
                if '%' in label and '2021' in label and '2022' in label:
                    pct_change_row1 = cell.row
                elif '%' in label and '2022' in label and '2023' in label:
                    pct_change_row2 = cell.row

        if pct_change_row1 is None or pct_change_row2 is None:
            print(f"FAIL: Component 2 — Percentage change rows not found (pct_row1={pct_change_row1}, pct_row2={pct_change_row2})")
        else:
            # Verify % Change 2021→2022 values (decimal form, e.g. 0.0833)
            pct1_correct = 0
            pct1_total = 6
            for col_idx in range(2, 8):  # B to G
                actual = ws.cell(row=pct_change_row1, column=col_idx).value
                expected = EXPECTED_PCT_2021_2022[col_idx - 2]
                # Allow small tolerance since values could be stored with varying precision
                if close_enough(actual, expected, tol=0.005):
                    pct1_correct += 1
                else:
                    print(f"  FAIL pct 2021->2022 col {col_idx}: expected {expected}, got {actual}")

            # Verify % Change 2022→2023 values
            pct2_correct = 0
            pct2_total = 6
            for col_idx in range(2, 8):  # B to G
                actual = ws.cell(row=pct_change_row2, column=col_idx).value
                expected = EXPECTED_PCT_2022_2023[col_idx - 2]
                if close_enough(actual, expected, tol=0.005):
                    pct2_correct += 1
                else:
                    print(f"  FAIL pct 2022->2023 col {col_idx}: expected {expected}, got {actual}")

            total_correct = pct1_correct + pct2_correct
            total_possible = pct1_total + pct2_total

            if total_correct == total_possible:
                print(f"PASS: Component 2 — Both percentage change rows correct ({total_correct}/{total_possible} values) (0.40 pts)")
                total_score += 0.40
            elif total_correct >= total_possible * 0.5:
                print(f"PARTIAL: Component 2 — {total_correct}/{total_possible} percentage change values correct (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 2 — Only {total_correct}/{total_possible} percentage change values correct")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------
    # Component 3: Ranking row present and correct (0.30 points)
    # Row labeled 'Rank ...' must exist with ranks 1-6 assigned by
    # descending average YoY growth rate across the two transitions
    # Expected ranks (B-G): 3, 6, 2, 5, 1, 4
    # -------------------------------------------------------------------
    try:
        rank_row = None
        for row in ws.iter_rows(min_col=1, max_col=1):
            cell = row[0]
            if cell.value:
                label = str(cell.value).lower()
                if 'rank' in label:
                    rank_row = cell.row
                    break

        if rank_row is None:
            print("FAIL: Component 3 — Ranking row not found (no 'Rank' label in column A)")
        else:
            rank_correct = 0
            rank_total = 6
            actual_ranks = []
            for col_idx in range(2, 8):  # B to G
                actual = ws.cell(row=rank_row, column=col_idx).value
                actual_ranks.append(actual)
                expected = EXPECTED_RANKS[col_idx - 2]
                if actual is not None and int(actual) == expected:
                    rank_correct += 1
                else:
                    print(f"  FAIL rank col {col_idx}: expected rank {expected}, got {actual}")

            print(f"  INFO: Actual ranks (B-G): {actual_ranks}")
            print(f"  INFO: Expected ranks (B-G): {EXPECTED_RANKS}")

            if rank_correct == rank_total:
                print(f"PASS: Component 3 — Ranking row correct ({rank_correct}/{rank_total} ranks correct) (0.30 pts)")
                total_score += 0.30
            elif rank_correct >= rank_total * 0.5:
                print(f"PARTIAL: Component 3 — {rank_correct}/{rank_total} ranks correct (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 3 — Only {rank_correct}/{rank_total} ranks correct")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in a given env
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
