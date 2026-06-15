"""
Initial Setup: Multi-app document follow instructions task
Task ID: osworld_multi_apps_doc_follow_instructions_010
Domain: libreoffice_calc + libreoffice_writer
Creates: analysis_instructions.odt, research_data.ods, paper_draft.odt in /home/user/Documents
"""

import os
import shlex
import subprocess
import sys
import time

# Install required dependencies before any imports that need them
subprocess.run([sys.executable, '-m', 'pip', 'install', 'odfpy', 'openpyxl', '-q'],
               capture_output=True)

WORKDIR = '/home/user'
DOCS_DIR = '/home/user/Documents'
TASK_ID = 'osworld_multi_apps_doc_follow_instructions_010'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_analysis_instructions():
    """Create the instruction document listing all 11 analysis steps."""
    import subprocess

    odt_path = os.path.join(DOCS_DIR, 'analysis_instructions.odt')

    # Build ODT using python-odf
    from odf.opendocument import OpenDocumentText
    from odf.text import H, P
    from odf.style import Style, TextProperties, ParagraphProperties

    doc = OpenDocumentText()

    # Add heading style
    h1_style = Style(name="Heading1Custom", family="paragraph")
    h1_style.addElement(TextProperties(fontsize="14pt", fontweight="bold"))
    doc.automaticstyles.addElement(h1_style)

    # Title
    title = H(outlinelevel=1)
    title.addText("Research Data Analysis Instructions")
    doc.text.addElement(title)

    # Intro
    intro = P()
    intro.addText("The following steps must be completed in order. Perform all Calc operations on research_data.ods, then update paper_draft.odt with the findings.")
    doc.text.addElement(intro)

    # Section header for Calc operations
    calc_header = H(outlinelevel=2)
    calc_header.addText("Calc Operations on research_data.ods")
    doc.text.addElement(calc_header)

    calc_steps = [
        "1. Remove duplicate rows based on the ID column. Keep the first occurrence of each duplicate.",
        "2. Fill blank cells in column C (Category) using forward-fill: each blank cell takes the value of the nearest non-blank cell above it.",
        "3. Normalize column D (Value) to the 0-1 range using min-max normalization: (x - min) / (max - min). Replace the original values in column D.",
        "4. Add quartile labels in column E (Quartile): Q1 for values in the lowest 25%, Q2 for 25-50%, Q3 for 50-75%, Q4 for the top 25%. Use the normalized D values for quartile assignment.",
        "5. Create a frequency table on a new sheet named 'FrequencyTable'. The table should count the number of occurrences for each unique Category value from column C.",
        "6. Generate a histogram chart on the main data sheet (Sheet1) showing the distribution of normalized D values. The chart should be a bar/column chart with the x-axis representing value bins.",
        "7. Run and record descriptive statistics for the normalized column D values in a new sheet named 'Statistics'. Include: mean, standard deviation (std), median, and interquartile range (IQR).",
    ]

    for step in calc_steps:
        p = P()
        p.addText(step)
        doc.text.addElement(p)

    # Section header for Writer insertions
    writer_header = H(outlinelevel=2)
    writer_header.addText("Writer Insertions in paper_draft.odt")
    doc.text.addElement(writer_header)

    writer_steps = [
        "8. Replace the placeholder [SAMPLE_SIZE] in paper_draft.odt with the actual number of data rows remaining after duplicate removal.",
        "9. Insert the descriptive statistics (mean, std, median, IQR) as a formatted table at the [STATS_TABLE] marker in paper_draft.odt. The table should have two columns: Statistic and Value.",
        "10. Update the chart figure reference in paper_draft.odt. Replace [CHART_REF] with 'Figure 1: Distribution of Normalized Values'.",
        "11. Update the word count in the abstract of paper_draft.odt. Replace [WORD_COUNT] with the actual approximate word count of the abstract section.",
    ]

    for step in writer_steps:
        p = P()
        p.addText(step)
        doc.text.addElement(p)

    doc.save(odt_path)
    print(f"Created: {odt_path}")


def create_research_data():
    """Create the research_data.ods spreadsheet with raw data including duplicates and blanks."""
    import openpyxl
    from openpyxl import Workbook

    ods_path = os.path.join(DOCS_DIR, 'research_data.ods')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # Headers: ID, Name, Category, Value, Quartile (E is empty - agent must fill)
    headers = ['ID', 'Name', 'Category', 'Value', 'Quartile']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Data rows - 20 rows total with:
    # - Row 8 is a duplicate of row 3 (same ID=103)
    # - Row 15 is a duplicate of row 7 (same ID=107)
    # - Column C (Category) has blanks at rows 4, 5, 9, 13, 16 (0-indexed from data start)
    # - Column D (Value) raw numerical values 0-100 range
    # After removing duplicates: 18 rows remain

    # Format: [ID, Name, Category, Value]
    # None means blank cell
    raw_data = [
        [101, 'Alice Novak',       'Biology',   45.2],   # row 2
        [102, 'Ben Okafor',        'Chemistry', 78.5],   # row 3
        [103, 'Carol Petersen',    'Biology',   32.1],   # row 4
        [104, 'David Quiroz',      None,        61.7],   # row 5  - blank C
        [105, 'Elena Rashid',      None,        88.4],   # row 6  - blank C
        [106, 'Felix Sato',        'Physics',   19.3],   # row 7
        [107, 'Grace Turner',      'Chemistry', 55.0],   # row 8
        [108, 'Hiro Ueda',         None,        42.6],   # row 9  - blank C
        [109, 'Irene Valdez',      'Biology',   73.8],   # row 10
        [110, 'James Wong',        'Physics',   91.2],   # row 11
        [111, 'Kira Xu',           'Chemistry', 28.9],   # row 12
        [112, 'Leo Yamamoto',      'Biology',   66.4],   # row 13
        [113, 'Maya Zafar',        None,        50.1],   # row 14 - blank C
        [114, 'Nate Abreu',        'Physics',   37.8],   # row 15
        [115, 'Olivia Becker',     'Chemistry', 82.3],   # row 16
        [103, 'Carol Petersen',    'Biology',   32.1],   # row 17 - DUPLICATE of row 4
        [116, 'Pablo Cruz',        None,        15.6],   # row 18 - blank C
        [117, 'Quinn Dolan',       'Physics',   70.2],   # row 19
        [107, 'Grace Turner',      'Chemistry', 55.0],   # row 20 - DUPLICATE of row 8
        [118, 'Raquel Escobar',    'Biology',   47.9],   # row 21
    ]

    for r, row_data in enumerate(raw_data, 2):
        ws.cell(row=r, column=1, value=row_data[0])  # ID
        ws.cell(row=r, column=2, value=row_data[1])  # Name
        if row_data[2] is not None:
            ws.cell(row=r, column=3, value=row_data[2])  # Category (some blanks)
        ws.cell(row=r, column=4, value=row_data[3])  # Value
        # Column E (Quartile) intentionally left blank

    # Save as xlsx but name it .ods for task authenticity
    # LibreOffice can open xlsx files saved with .ods extension
    # Actually we should use xlsx format internally, just save with proper name
    # Using .xlsx format but saving as .ods for compatibility
    wb.save(ods_path)
    print(f"Created: {ods_path}")
    return raw_data


def create_paper_draft():
    """Create the paper draft ODT with placeholders for the agent to fill."""
    from odf.opendocument import OpenDocumentText
    from odf.text import H, P
    from odf.style import Style, TextProperties

    odt_path = os.path.join(DOCS_DIR, 'paper_draft.odt')

    doc = OpenDocumentText()

    # Title
    title = H(outlinelevel=1)
    title.addText("Exploratory Analysis of Multi-Domain Research Data")
    doc.text.addElement(title)

    # Authors
    authors = P()
    authors.addText("Dr. Sarah Chen, Prof. Marcus Johnson, Dr. Elena Rashid")
    doc.text.addElement(authors)

    # Abstract section
    abstract_header = H(outlinelevel=2)
    abstract_header.addText("Abstract")
    doc.text.addElement(abstract_header)

    abstract = P()
    abstract.addText(
        "This paper presents an exploratory analysis of a multi-domain research dataset comprising "
        "[SAMPLE_SIZE] observations across three scientific disciplines: Biology, Chemistry, and Physics. "
        "Using standard statistical techniques, we perform data cleaning including duplicate removal and "
        "missing value imputation via forward-fill methodology. The cleaned dataset undergoes min-max "
        "normalization to facilitate cross-domain comparisons. We report key descriptive statistics "
        "and visualize the distribution of normalized values. The word count of this abstract is approximately "
        "[WORD_COUNT] words. The dataset and analysis scripts are available upon request."
    )
    doc.text.addElement(abstract)

    # Introduction
    intro_header = H(outlinelevel=2)
    intro_header.addText("1. Introduction")
    doc.text.addElement(intro_header)

    intro1 = P()
    intro1.addText(
        "The integration of data from multiple scientific domains presents unique analytical challenges. "
        "Raw datasets frequently contain duplicated entries, missing values, and heterogeneous value ranges "
        "that must be addressed before meaningful analysis can proceed."
    )
    doc.text.addElement(intro1)

    intro2 = P()
    intro2.addText(
        "The present study examines a dataset of [SAMPLE_SIZE] research observations collected across "
        "Biology, Chemistry, and Physics departments over the 2024-2025 academic year. Each observation "
        "includes a unique identifier, researcher name, disciplinary category, and a continuous measurement value."
    )
    doc.text.addElement(intro2)

    # Methods
    methods_header = H(outlinelevel=2)
    methods_header.addText("2. Methods")
    doc.text.addElement(methods_header)

    methods1 = P()
    methods1.addText(
        "Data preprocessing was performed in three stages. First, duplicate records were identified "
        "using the ID column and removed, retaining only the first occurrence of each duplicated entry. "
        "Second, missing category values were imputed using forward-fill, where each blank cell inherits "
        "the value from the nearest preceding non-blank cell. Third, the continuous measurement values "
        "were normalized to the [0, 1] range using min-max normalization."
    )
    doc.text.addElement(methods1)

    methods2 = P()
    methods2.addText(
        "Following normalization, observations were assigned to quartile groups (Q1 through Q4) based on "
        "their normalized values. A frequency distribution analysis was conducted to examine the distribution "
        "of observations across scientific disciplines."
    )
    doc.text.addElement(methods2)

    # Results
    results_header = H(outlinelevel=2)
    results_header.addText("3. Results")
    doc.text.addElement(results_header)

    results1 = P()
    results1.addText(
        "After data cleaning, the dataset contained [SAMPLE_SIZE] unique observations. "
        "The distribution of normalized measurement values is shown in [CHART_REF]."
    )
    doc.text.addElement(results1)

    results2 = P()
    results2.addText(
        "Descriptive statistics for the normalized values are summarized below:"
    )
    doc.text.addElement(results2)

    stats_marker = P()
    stats_marker.addText("[STATS_TABLE]")
    doc.text.addElement(stats_marker)

    results3 = P()
    results3.addText(
        "The frequency distribution analysis revealed that Biology had the highest representation "
        "in the dataset, followed by Chemistry and Physics. The quartile distribution showed an "
        "approximately uniform spread of observations across all four quartiles."
    )
    doc.text.addElement(results3)

    # Discussion
    disc_header = H(outlinelevel=2)
    disc_header.addText("4. Discussion")
    doc.text.addElement(disc_header)

    disc1 = P()
    disc1.addText(
        "The results demonstrate that standard preprocessing techniques are effective for handling "
        "the data quality issues present in this multi-domain dataset. The normalization step enables "
        "direct comparison of measurement values across disciplines, which would otherwise be confounded "
        "by domain-specific measurement scales."
    )
    doc.text.addElement(disc1)

    disc2 = P()
    disc2.addText(
        "The quartile analysis suggests that measurement values are distributed across a wide range "
        "within each scientific domain, indicating natural variation in the research outcomes measured."
    )
    doc.text.addElement(disc2)

    # Conclusion
    conc_header = H(outlinelevel=2)
    conc_header.addText("5. Conclusion")
    doc.text.addElement(conc_header)

    conc = P()
    conc.addText(
        "This study presents a systematic approach to cleaning and analyzing multi-domain research data. "
        "The preprocessing pipeline and statistical analysis framework described here can be applied "
        "to similar datasets in future research. The cleaned dataset with [SAMPLE_SIZE] observations "
        "provides a foundation for further domain-specific analyses."
    )
    doc.text.addElement(conc)

    doc.save(odt_path)
    print(f"Created: {odt_path}")


def create_initial():
    """Main function to create all initial files."""
    # Ensure Documents directory exists
    os.makedirs(DOCS_DIR, exist_ok=True)

    # Install dependencies first
    subprocess.run(['pip3', 'install', 'odfpy', 'openpyxl', '-q'], capture_output=True)

    create_analysis_instructions()
    create_research_data()
    create_paper_draft()

    print(f"\nAll initial files created in {DOCS_DIR}")

    # GUI-ready startup: open analysis_instructions.odt first for the agent to read
    instructions_path = os.path.join(DOCS_DIR, 'analysis_instructions.odt')
    launch_gui(f'libreoffice --writer "{instructions_path}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Writer with analysis_instructions.odt')


create_initial()
