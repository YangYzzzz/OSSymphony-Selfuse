"""
Reward Script: Multi-app research data analysis task
Task ID: osworld_multi_apps_doc_follow_instructions_010
Domain: libreoffice_calc + libreoffice_writer
Scoring:
  - research_data.ods: duplicate removal (rows == 18), normalization, quartiles, FrequencyTable sheet, Statistics sheet, chart
  - paper_draft.odt: [SAMPLE_SIZE] replaced, stats table inserted, [CHART_REF] replaced, [WORD_COUNT] replaced
"""

import os
import io

WORKDIR = '/home/user/Documents'
TASK_ID = 'osworld_multi_apps_doc_follow_instructions_010'

ODS_PATH = f'{WORKDIR}/research_data.ods'
ODT_PATH = f'{WORKDIR}/paper_draft.odt'


def verify_task():
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # =========================================================================
    # PART 1: research_data.ods verification
    # =========================================================================

    # Load the ODS file (it is actually xlsx format internally)
    try:
        import openpyxl
        with open(ODS_PATH, 'rb') as f:
            data = f.read()
        wb = openpyxl.load_workbook(io.BytesIO(data))
    except Exception as e:
        print(f"CRITICAL: Cannot load {ODS_PATH}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # -------------------------------------------------------------------------
    # Component 1: Duplicate removal and forward-fill — row count should be 18
    # After removing 2 duplicate rows (ID 103 and 107), count = 20 - 2 = 18.
    # Also verify column C (Category) has no blank cells in data rows (forward fill).
    # (0.20 points)
    # -------------------------------------------------------------------------
    try:
        if 'Sheet1' not in wb.sheetnames:
            print("FAIL: Component 1 — Sheet1 not found")
        else:
            ws = wb['Sheet1']
            # Count data rows (skip header row 1)
            data_rows = ws.max_row - 1  # subtract header
            if data_rows == 18:
                print(f"PASS: Component 1a — Row count is 18 after duplicate removal (data_rows={data_rows})")
                total_score += 0.10
            else:
                print(f"FAIL: Component 1a — Expected 18 data rows, found {data_rows}")

            # Check column C has no blank cells in data rows
            blank_c = 0
            for i in range(2, ws.max_row + 1):
                val = ws.cell(row=i, column=3).value
                if val is None or str(val).strip() == '':
                    blank_c += 1
            if blank_c == 0:
                print(f"PASS: Component 1b — Column C has no blank cells (forward-fill applied)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 1b — Column C has {blank_c} blank cells, expected 0")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Column D values normalized to 0-1 range
    # All values should be in [0.0, 1.0], min should be 0, max should be 1.
    # (0.15 points)
    # -------------------------------------------------------------------------
    try:
        if 'Sheet1' in wb.sheetnames:
            ws = wb['Sheet1']
            d_values = []
            for i in range(2, ws.max_row + 1):
                val = ws.cell(row=i, column=4).value
                if val is not None:
                    try:
                        d_values.append(float(val))
                    except (ValueError, TypeError):
                        pass

            if d_values:
                all_in_range = all(0.0 <= v <= 1.0 for v in d_values)
                min_val = min(d_values)
                max_val = max(d_values)
                # Min should be close to 0, max should be close to 1
                min_ok = abs(min_val - 0.0) < 0.001
                max_ok = abs(max_val - 1.0) < 0.001

                if all_in_range and min_ok and max_ok:
                    print(f"PASS: Component 2 — Column D normalized to [0,1], min={min_val:.4f}, max={max_val:.4f}")
                    total_score += 0.15
                elif all_in_range:
                    print(f"FAIL: Component 2 — Values in [0,1] but min={min_val:.4f} (expected ~0) or max={max_val:.4f} (expected ~1)")
                else:
                    print(f"FAIL: Component 2 — Values not fully normalized. Min={min_val:.4f}, Max={max_val:.4f}")
            else:
                print("FAIL: Component 2 — No numeric values in column D")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Column E (Quartile) filled with Q1/Q2/Q3/Q4 labels
    # (0.10 points)
    # -------------------------------------------------------------------------
    try:
        if 'Sheet1' in wb.sheetnames:
            ws = wb['Sheet1']
            valid_labels = {'Q1', 'Q2', 'Q3', 'Q4'}
            quartile_vals = []
            for i in range(2, ws.max_row + 1):
                val = ws.cell(row=i, column=5).value
                quartile_vals.append(val)

            non_null = [v for v in quartile_vals if v is not None]
            valid_count = sum(1 for v in non_null if str(v) in valid_labels)
            total_count = len(quartile_vals)

            if len(non_null) == total_count and valid_count == total_count:
                print(f"PASS: Component 3 — Column E has quartile labels (Q1/Q2/Q3/Q4) for all {total_count} rows")
                total_score += 0.10
            else:
                print(f"FAIL: Component 3 — Column E: {valid_count}/{total_count} valid quartile labels, non-null={len(non_null)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: FrequencyTable sheet exists with correct category counts
    # Expected: Biology=8, Chemistry=6, Physics=4 (total=18 rows)
    # (0.10 points)
    # -------------------------------------------------------------------------
    try:
        if 'FrequencyTable' not in wb.sheetnames:
            print("FAIL: Component 4 — Sheet 'FrequencyTable' not found")
        else:
            ws_freq = wb['FrequencyTable']
            # Read the frequency table
            freq_data = {}
            for i in range(2, ws_freq.max_row + 1):
                cat = ws_freq.cell(row=i, column=1).value
                cnt = ws_freq.cell(row=i, column=2).value
                if cat and cnt is not None:
                    freq_data[str(cat)] = int(cnt)

            expected_freq = {'Biology': 8, 'Chemistry': 6, 'Physics': 4}
            if freq_data == expected_freq:
                print(f"PASS: Component 4 — FrequencyTable has correct counts: {freq_data}")
                total_score += 0.10
            else:
                print(f"FAIL: Component 4 — FrequencyTable has {freq_data}, expected {expected_freq}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -------------------------------------------------------------------------
    # Component 5: Statistics sheet exists with mean, std, median, IQR values
    # Expected values: Mean~0.519, Std~0.307, Median~0.489, IQR~0.476
    # (0.10 points)
    # -------------------------------------------------------------------------
    try:
        if 'Statistics' not in wb.sheetnames:
            print("FAIL: Component 5 — Sheet 'Statistics' not found")
        else:
            ws_stats = wb['Statistics']
            stats_data = {}
            for i in range(2, ws_stats.max_row + 1):
                stat_name = ws_stats.cell(row=i, column=1).value
                stat_val = ws_stats.cell(row=i, column=2).value
                if stat_name and stat_val is not None:
                    stats_data[str(stat_name)] = float(stat_val)

            # Check all four statistics are present
            required_stats = ['Mean', 'Std', 'Median', 'IQR']
            present = [s for s in required_stats if s in stats_data]

            if len(present) == 4:
                # Verify approximate values
                mean_ok = abs(stats_data.get('Mean', -1) - 0.518959) < 0.01
                std_ok = abs(stats_data.get('Std', -1) - 0.306624) < 0.01
                median_ok = abs(stats_data.get('Median', -1) - 0.488757) < 0.01
                iqr_ok = abs(stats_data.get('IQR', -1) - 0.47619) < 0.01

                if mean_ok and std_ok and median_ok and iqr_ok:
                    print(f"PASS: Component 5 — Statistics sheet has correct values: {stats_data}")
                    total_score += 0.10
                else:
                    print(f"FAIL: Component 5 — Statistics values incorrect. mean_ok={mean_ok}, std_ok={std_ok}, median_ok={median_ok}, iqr_ok={iqr_ok}. Values: {stats_data}")
            else:
                print(f"FAIL: Component 5 — Statistics sheet missing stats. Found: {list(stats_data.keys())}, expected: {required_stats}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # -------------------------------------------------------------------------
    # Component 6: Histogram/bar chart exists in Sheet1
    # (0.05 points)
    # -------------------------------------------------------------------------
    try:
        if 'Sheet1' in wb.sheetnames:
            ws = wb['Sheet1']
            charts = ws._charts
            if len(charts) >= 1:
                chart = charts[0]
                chart_type = type(chart).__name__
                print(f"PASS: Component 6 — Chart found in Sheet1 (type={chart_type})")
                total_score += 0.05
            else:
                print("FAIL: Component 6 — No chart found in Sheet1")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # =========================================================================
    # PART 2: paper_draft.odt verification
    # =========================================================================

    try:
        from odf.opendocument import load as odf_load
        from odf.table import Table, TableRow, TableCell
        from odf.text import P
        odt_doc = odf_load(ODT_PATH)
    except Exception as e:
        print(f"CRITICAL: Cannot load {ODT_PATH}: {e}")
        print(f"\nScore: {total_score}/1.0")
        final_score = min(total_score, 1.0)
        print(f"REWARD: {final_score}")
        return final_score

    # Helper to get all text from an ODT element
    def get_all_text(node):
        text = ''
        if hasattr(node, 'nodeType') and node.nodeType == 3:
            return node.data if hasattr(node, 'data') else ''
        if hasattr(node, 'childNodes'):
            for child in node.childNodes:
                text += get_all_text(child)
        return text

    # Get full text of document
    full_text = get_all_text(odt_doc.text)

    # -------------------------------------------------------------------------
    # Component 7: [SAMPLE_SIZE] replaced with 18 in paper_draft.odt
    # (0.10 points)
    # -------------------------------------------------------------------------
    try:
        if '[SAMPLE_SIZE]' in full_text:
            print("FAIL: Component 7 — [SAMPLE_SIZE] placeholder still present in paper_draft.odt")
        elif '18' in full_text:
            # Verify "18" appears where SAMPLE_SIZE was (in abstract/intro/conclusion)
            print("PASS: Component 7 — [SAMPLE_SIZE] replaced with '18' in paper_draft.odt")
            total_score += 0.10
        else:
            print("FAIL: Component 7 — [SAMPLE_SIZE] replaced but '18' not found in text")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    # -------------------------------------------------------------------------
    # Component 8: [STATS_TABLE] replaced with actual statistics table
    # The table should have 5 rows (header + 4 stats: Mean, Std, Median, IQR)
    # (0.10 points)
    # -------------------------------------------------------------------------
    try:
        if '[STATS_TABLE]' in full_text:
            print("FAIL: Component 8 — [STATS_TABLE] placeholder still present in paper_draft.odt")
        else:
            tables = odt_doc.getElementsByType(Table)
            if len(tables) >= 1:
                # Check the table has expected structure (at least 5 rows with stats data)
                table = tables[0]
                rows = table.getElementsByType(TableRow)
                if len(rows) >= 5:
                    # Check header row
                    header_cells = rows[0].getElementsByType(TableCell)
                    header_texts = [get_all_text(c).strip() for c in header_cells]
                    has_statistic = any('Statistic' in t for t in header_texts)
                    has_value = any('Value' in t for t in header_texts)
                    if has_statistic and has_value:
                        print(f"PASS: Component 8 — Stats table inserted with {len(rows)} rows and correct headers")
                        total_score += 0.10
                    else:
                        print(f"FAIL: Component 8 — Table found but headers incorrect: {header_texts}")
                else:
                    print(f"FAIL: Component 8 — Table found but only {len(rows)} rows, expected >=5")
            else:
                print("FAIL: Component 8 — No table found in paper_draft.odt and [STATS_TABLE] removed")
    except Exception as e:
        print(f"ERROR: Component 8 — {e}")

    # -------------------------------------------------------------------------
    # Component 9: [CHART_REF] replaced with figure reference text
    # Expected: "Figure 1: Distribution of Normalized Values"
    # (0.05 points)
    # -------------------------------------------------------------------------
    try:
        if '[CHART_REF]' in full_text:
            print("FAIL: Component 9 — [CHART_REF] placeholder still present in paper_draft.odt")
        elif 'Figure 1' in full_text or 'figure 1' in full_text.lower():
            print("PASS: Component 9 — [CHART_REF] replaced with figure reference text")
            total_score += 0.05
        else:
            print(f"FAIL: Component 9 — [CHART_REF] removed but 'Figure 1' not found in text")
    except Exception as e:
        print(f"ERROR: Component 9 — {e}")

    # -------------------------------------------------------------------------
    # Component 10: [WORD_COUNT] replaced with a numeric value
    # (0.05 points)
    # -------------------------------------------------------------------------
    try:
        if '[WORD_COUNT]' in full_text:
            print("FAIL: Component 10 — [WORD_COUNT] placeholder still present in paper_draft.odt")
        else:
            # Look for a pattern like "approximately NNN words" near word count
            import re
            # Find digits near word count context
            wc_pattern = re.search(r'approximately\s+(\d+)\s+words', full_text)
            if wc_pattern:
                wc_value = int(wc_pattern.group(1))
                print(f"PASS: Component 10 — [WORD_COUNT] replaced with numeric value {wc_value}")
                total_score += 0.05
            else:
                # Might be present but without exactly "approximately X words" format
                # Check that text doesn't have [WORD_COUNT] and has some number
                print("FAIL: Component 10 — [WORD_COUNT] removed but numeric word count pattern not found")
    except Exception as e:
        print(f"ERROR: Component 10 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


if __name__ == '__main__':
    # Check files exist before running
    if not os.path.exists(ODS_PATH):
        print(f"File not found: {ODS_PATH}")
        print("REWARD: 0.0")
    elif not os.path.exists(ODT_PATH):
        print(f"File not found: {ODT_PATH}")
        print("REWARD: 0.0")
    else:
        verify_task()
