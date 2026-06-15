"""
Initial Setup: Apply validation to cell B2 with dropdown list
Task ID: calc_nrv_083
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_083'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Sheet1"

    # Headers
    ws.cell(row=1, column=1, value="Employee")
    ws.cell(row=1, column=2, value="Attending?")

    # Realistic employee data in column A; column B left empty (task is to add validation)
    employees = [
        "Sarah Chen",
        "Marcus Johnson",
        "Priya Patel",
        "David Kim",
        "Elena Rodriguez",
        "James Thompson",
        "Aisha Williams",
        "Robert Garcia",
        "Mei Lin",
        "Carlos Fernandez",
        "Jessica Adams",
        "Tomoko Yamada",
    ]
    for r, name in enumerate(employees, 2):
        ws.cell(row=r, column=1, value=name)
        # B column intentionally left empty -- no validation, no values

    # Adjust column widths for readability
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
