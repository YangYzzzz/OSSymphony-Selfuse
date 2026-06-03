"""
Initial Setup: Fleet Maintenance Schedule
Task ID: calc_ops_fleet_maintenance_schedule_030
Domain: libreoffice_calc

Creates a 'MaintenanceSchedule' sheet with 10 vehicles.
Columns A-D filled with realistic data. Columns E-J left empty for agent to populate.
"""

import os
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import date

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_fleet_maintenance_schedule_030'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'MaintenanceSchedule'

    # --- Headers (Row 1) ---
    headers = [
        'Vehicle ID',
        'Last Service Date',
        'Last Service Odometer',
        'Current Odometer',
        'Km Since Service',
        'Next Service Date by Months',
        'Next Service Date by Km',
        'Next Service Due',
        'Days Until Service',
        'Alert'
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.row_dimensions[1].height = 30

    # Column widths
    col_widths = [14, 20, 24, 18, 18, 28, 24, 18, 18, 14]
    for idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # --- Vehicle data (rows 2-11) ---
    # Mix of SERVICE DUE and OK vehicles relative to 2025-03-04
    # (Vehicle ID, Last Service Date, Last Service Odometer, Current Odometer)
    # SERVICE DUE: TRK-001, TRK-002, TRK-003, TRK-004, TRK-007, TRK-010
    # OK:          TRK-005, TRK-006, TRK-008, TRK-009
    vehicles = [
        ('TRK-001', date(2025, 1, 10), 145000, 154500),   # km_since=9500, next_km=Jan11 (overdue)
        ('TRK-002', date(2024, 12, 1),  203400, 205000),   # next_months=Mar1 (overdue)
        ('TRK-003', date(2025, 2, 20),   87600,  94100),   # km_since=6500, next_km=Feb27 (overdue)
        ('TRK-004', date(2025, 2, 24),  312100, 315600),   # km_since=3500, next_km=Mar9 (5 days)
        ('TRK-005', date(2025, 3, 1),    56800,  57800),   # km_since=1000, next_km=Mar19 (15 days OK)
        ('TRK-006', date(2025, 2, 28),  178500, 179000),   # km_since=500,  next_km=Mar19 (15 days OK)
        ('TRK-007', date(2025, 2, 25),  224300, 224800),   # km_since=500,  next_km=Mar16 (12 days)
        ('TRK-008', date(2025, 3, 2),   441000, 441500),   # km_since=500,  next_km=Mar21 (17 days OK)
        ('TRK-009', date(2025, 3, 1),   130700, 131200),   # km_since=500,  next_km=Mar20 (16 days OK)
        ('TRK-010', date(2025, 2, 14),  289200, 296800),   # km_since=7600, next_km=Feb18 (overdue)
    ]

    for r, (vid, last_svc, last_odo, curr_odo) in enumerate(vehicles, 2):
        ws.cell(row=r, column=1, value=vid)
        date_cell = ws.cell(row=r, column=2, value=last_svc)
        date_cell.number_format = 'yyyy-mm-dd'
        ws.cell(row=r, column=3, value=last_odo)
        ws.cell(row=r, column=4, value=curr_odo)
        # Columns E-J intentionally left empty (agent must fill them)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
