"""
Reward Script: Build a preventive maintenance schedule for truck fleet
Task ID: calc_ops_fleet_maintenance_schedule_030
Domain: libreoffice_calc
Scoring:
  - Component 1: E2:E11 km-since-service formulas (=D-C)           0.20 pts
  - Component 2: F2:F11 next-service-by-months formulas (=B+90)    0.20 pts
  - Component 3: G2:G11 next-service-by-km formulas                0.20 pts
  - Component 4: H2:H11 MIN formula (=MIN(F,G))                    0.15 pts
  - Component 5: I2:I11 days-until-service formulas (=H-TODAY())   0.10 pts
  - Component 6: J2:J11 IF alert formulas                          0.05 pts
  - Component 7: Red fill on SERVICE DUE vehicle rows               0.10 pts
  Total: 1.0
"""

import os
import openpyxl
from openpyxl.utils import get_column_letter

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_fleet_maintenance_schedule_030'

DATA_ROWS = list(range(2, 12))  # rows 2-11 (10 vehicles)


def normalize_formula(f):
    """Normalize formula string for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '').strip()


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet exists
    if 'MaintenanceSchedule' not in wb.sheetnames:
        print("CRITICAL: Sheet 'MaintenanceSchedule' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['MaintenanceSchedule']

    # Component 1: E2:E11 — Km Since Service formula =D{row}-C{row} (0.20 pts)
    # FAILS on initial (cells are empty), PASSES on golden (formulas present)
    try:
        e_pass = 0
        e_fails = []
        for row in DATA_ROWS:
            expected = f'=D{row}-C{row}'
            val = ws.cell(row=row, column=5).value
            if normalize_formula(val) == normalize_formula(expected):
                e_pass += 1
            else:
                e_fails.append(f'E{row}: got {repr(val)}')
        if e_pass == 10:
            print(f"PASS: Component 1 — All 10 E-column km-since-service formulas correct (0.20 pts)")
            total_score += 0.20
        elif e_pass >= 5:
            partial = round(0.20 * e_pass / 10, 4)
            print(f"PARTIAL: Component 1 — {e_pass}/10 E-column formulas correct ({partial} pts): {e_fails[:3]}")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {e_pass}/10 E-column formulas correct: {e_fails[:3]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: F2:F11 — Next Service by Months formula =B{row}+90 (0.20 pts)
    # FAILS on initial (cells are empty), PASSES on golden (formulas present)
    try:
        f_pass = 0
        f_fails = []
        for row in DATA_ROWS:
            expected = f'=B{row}+90'
            val = ws.cell(row=row, column=6).value
            if normalize_formula(val) == normalize_formula(expected):
                f_pass += 1
            else:
                f_fails.append(f'F{row}: got {repr(val)}')
        if f_pass == 10:
            print(f"PASS: Component 2 — All 10 F-column next-service-by-months formulas correct (0.20 pts)")
            total_score += 0.20
        elif f_pass >= 5:
            partial = round(0.20 * f_pass / 10, 4)
            print(f"PARTIAL: Component 2 — {f_pass}/10 F-column formulas correct ({partial} pts): {f_fails[:3]}")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {f_pass}/10 F-column formulas correct: {f_fails[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: G2:G11 — Next Service by Km formula =B{row}+((10000-E{row})/500) (0.20 pts)
    # FAILS on initial (cells are empty), PASSES on golden (formulas present)
    try:
        g_pass = 0
        g_fails = []
        for row in DATA_ROWS:
            expected = f'=B{row}+((10000-E{row})/500)'
            val = ws.cell(row=row, column=7).value
            if normalize_formula(val) == normalize_formula(expected):
                g_pass += 1
            else:
                g_fails.append(f'G{row}: got {repr(val)}')
        if g_pass == 10:
            print(f"PASS: Component 3 — All 10 G-column next-service-by-km formulas correct (0.20 pts)")
            total_score += 0.20
        elif g_pass >= 5:
            partial = round(0.20 * g_pass / 10, 4)
            print(f"PARTIAL: Component 3 — {g_pass}/10 G-column formulas correct ({partial} pts): {g_fails[:3]}")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {g_pass}/10 G-column formulas correct: {g_fails[:3]}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: H2:H11 — Next Service Due MIN formula =MIN(F{row},G{row}) (0.15 pts)
    # FAILS on initial (cells are empty), PASSES on golden (formulas present)
    try:
        h_pass = 0
        h_fails = []
        for row in DATA_ROWS:
            expected = f'=MIN(F{row},G{row})'
            val = ws.cell(row=row, column=8).value
            if normalize_formula(val) == normalize_formula(expected):
                h_pass += 1
            else:
                h_fails.append(f'H{row}: got {repr(val)}')
        if h_pass == 10:
            print(f"PASS: Component 4 — All 10 H-column MIN-next-service formulas correct (0.15 pts)")
            total_score += 0.15
        elif h_pass >= 5:
            partial = round(0.15 * h_pass / 10, 4)
            print(f"PARTIAL: Component 4 — {h_pass}/10 H-column formulas correct ({partial} pts): {h_fails[:3]}")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {h_pass}/10 H-column formulas correct: {h_fails[:3]}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: I2:I11 — Days Until Service formula =H{row}-TODAY() (0.10 pts)
    # FAILS on initial (cells are empty), PASSES on golden (formulas present)
    try:
        i_pass = 0
        i_fails = []
        for row in DATA_ROWS:
            expected = f'=H{row}-TODAY()'
            val = ws.cell(row=row, column=9).value
            if normalize_formula(val) == normalize_formula(expected):
                i_pass += 1
            else:
                i_fails.append(f'I{row}: got {repr(val)}')
        if i_pass == 10:
            print(f"PASS: Component 5 — All 10 I-column days-until-service formulas correct (0.10 pts)")
            total_score += 0.10
        elif i_pass >= 5:
            partial = round(0.10 * i_pass / 10, 4)
            print(f"PARTIAL: Component 5 — {i_pass}/10 I-column formulas correct ({partial} pts): {i_fails[:3]}")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — Only {i_pass}/10 I-column formulas correct: {i_fails[:3]}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: J2:J11 — Alert IF formula =IF(I{row}<=14,"SERVICE DUE","OK") (0.05 pts)
    # FAILS on initial (cells are empty), PASSES on golden (formulas present)
    try:
        j_pass = 0
        j_fails = []
        for row in DATA_ROWS:
            expected = f'=IF(I{row}<=14,"SERVICE DUE","OK")'
            val = ws.cell(row=row, column=10).value
            if normalize_formula(val) == normalize_formula(expected):
                j_pass += 1
            else:
                j_fails.append(f'J{row}: got {repr(val)}')
        if j_pass == 10:
            print(f"PASS: Component 6 — All 10 J-column alert-IF formulas correct (0.05 pts)")
            total_score += 0.05
        elif j_pass >= 5:
            partial = round(0.05 * j_pass / 10, 4)
            print(f"PARTIAL: Component 6 — {j_pass}/10 J-column formulas correct ({partial} pts): {j_fails[:3]}")
            total_score += partial
        else:
            print(f"FAIL: Component 6 — Only {j_pass}/10 J-column formulas correct: {j_fails[:3]}")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: Red fill (FFFF0000 solid) on SERVICE DUE vehicle rows (0.10 pts)
    # The golden file has red fill applied across all 10 columns for service-due rows.
    # FAILS on initial (no fill on any row), PASSES on golden (red fill on service-due rows).
    try:
        red_rows = []
        non_red_rows = []
        for row in DATA_ROWS:
            cell_a = ws.cell(row=row, column=1)
            try:
                fg = cell_a.fill.fgColor.rgb
                pt = cell_a.fill.patternType
                if fg == 'FFFF0000' and pt == 'solid':
                    red_rows.append(row)
                else:
                    non_red_rows.append(row)
            except Exception:
                non_red_rows.append(row)

        if len(red_rows) == 0:
            print("FAIL: Component 7 — No rows have red fill (FFFF0000 solid)")
        else:
            # Verify red fill covers all 10 columns on red rows
            cols_mismatch_count = 0
            for row in red_rows:
                for col in range(1, 11):
                    cell = ws.cell(row=row, column=col)
                    try:
                        fg = cell.fill.fgColor.rgb
                        pt = cell.fill.patternType
                        if not (fg == 'FFFF0000' and pt == 'solid'):
                            cols_mismatch_count += 1
                    except Exception:
                        cols_mismatch_count += 1

            # Verify non-highlighted rows have no spurious red fill
            spurious_count = 0
            for row in non_red_rows:
                cell_a = ws.cell(row=row, column=1)
                try:
                    fg = cell_a.fill.fgColor.rgb
                    pt = cell_a.fill.patternType
                    if fg == 'FFFF0000' and pt == 'solid':
                        spurious_count += 1
                except Exception:
                    pass

            comp7_full = (cols_mismatch_count == 0 and spurious_count == 0)
            comp7_partial = (not comp7_full) and (len(red_rows) > 0)
            if comp7_full:
                print(f"PASS: Component 7 — Red fill on {len(red_rows)} SERVICE DUE rows "
                      f"{red_rows}, no spurious highlights (0.10 pts)")
                total_score += 0.10
            elif comp7_partial:
                total_score += 0.05
                print(f"PARTIAL: Component 7 — Red fill partially correct: "
                      f"{len(red_rows)} red rows, {cols_mismatch_count} col mismatches, "
                      f"{spurious_count} spurious rows (0.05 pts)")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
