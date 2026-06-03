"""
Initial Setup: Insert hyperlink in Links sheet pointing to Archive sheet
Task ID: calc_gg3_011
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_011'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Links ---
    ws_links = wb.active
    ws_links.title = 'Links'

    # Headers
    headers_links = ['Navigation', 'Description', 'Last Updated']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    for col, h in enumerate(headers_links, 1):
        cell = ws_links.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # A1 must contain plain text placeholder (NO hyperlink)
    ws_links['A1'] = 'ARCHIVE LINK PLACEHOLDER'
    ws_links['A1'].font = Font(name='Calibri', size=11, bold=False)
    ws_links['A1'].fill = PatternFill()  # no fill, overriding header fill

    # Other navigation entries
    nav_data = [
        ['ARCHIVE LINK PLACEHOLDER', 'Link to archived historical records', '2025-11-01'],
        ['Summary Overview', 'Overview of key performance indicators', '2025-11-15'],
        ['Data Entry Portal', 'Raw data input section', '2025-10-28'],
        ['Budget Tracker', 'Annual budget monitoring dashboard', '2025-12-01'],
        ['Team Directory', 'Employee contact information', '2025-09-20'],
        ['Project Milestones', 'Timeline of key project deliverables', '2025-11-10'],
    ]
    for r, row_data in enumerate(nav_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_links.cell(row=r, column=c, value=val)

    # Overwrite row 2 col A with the placeholder (this is the actual A1 content area)
    # Actually A1 is row 1, let me fix the structure:
    # Row 1 = headers, but A1 is supposed to have "ARCHIVE LINK PLACEHOLDER"
    # The task says cell A1 contains the plain text. Let me restructure:
    # A1 = "ARCHIVE LINK PLACEHOLDER" (the cell the task targets)
    # Row 2+ = other content

    # Reset - A1 is the target cell with placeholder text
    ws_links['A1'] = 'ARCHIVE LINK PLACEHOLDER'
    ws_links['A1'].font = Font(name='Calibri', size=11)
    ws_links['A1'].fill = PatternFill()

    # Row 1 other cells as descriptive headers
    ws_links['B1'] = 'Description'
    ws_links['B1'].font = header_font
    ws_links['B1'].fill = header_fill
    ws_links['C1'] = 'Last Updated'
    ws_links['C1'].font = header_font
    ws_links['C1'].fill = header_fill

    # Navigation entries from row 2
    for r, row_data in enumerate(nav_data[1:], 2):  # skip first since A1 is special
        for c, val in enumerate(row_data, 1):
            ws_links.cell(row=r, column=c, value=val)

    # Column widths
    ws_links.column_dimensions['A'].width = 30
    ws_links.column_dimensions['B'].width = 45
    ws_links.column_dimensions['C'].width = 18

    # --- Sheet 2: Archive ---
    ws_archive = wb.create_sheet('Archive')

    archive_headers = ['Date', 'Category', 'Description', 'Amount', 'Status']
    for col, h in enumerate(archive_headers, 1):
        cell = ws_archive.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    archive_data = [
        ['2024-01-15', 'Operations', 'Server migration Phase 1', 12500.00, 'Completed'],
        ['2024-02-03', 'Marketing', 'Q1 campaign launch', 8750.00, 'Completed'],
        ['2024-02-28', 'Engineering', 'Database optimization project', 15200.00, 'Completed'],
        ['2024-03-10', 'HR', 'Annual employee survey rollout', 3400.00, 'Completed'],
        ['2024-04-05', 'Finance', 'Tax filing preparation', 6800.00, 'Completed'],
        ['2024-04-22', 'Operations', 'Office relocation planning', 22000.00, 'Completed'],
        ['2024-05-15', 'Marketing', 'Website redesign Phase 2', 18500.00, 'Completed'],
        ['2024-06-01', 'Engineering', 'API gateway upgrade', 9300.00, 'Completed'],
        ['2024-06-20', 'HR', 'New hire onboarding program', 4200.00, 'Completed'],
        ['2024-07-12', 'Finance', 'Mid-year budget review', 1500.00, 'Completed'],
        ['2024-08-08', 'Operations', 'Vendor contract renegotiation', 0.00, 'Completed'],
        ['2024-09-01', 'Marketing', 'Product launch event', 35000.00, 'Completed'],
        ['2024-10-15', 'Engineering', 'Security audit remediation', 11000.00, 'In Progress'],
        ['2024-11-01', 'HR', 'Benefits enrollment period', 2800.00, 'Pending'],
    ]
    for r, row_data in enumerate(archive_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_archive.cell(row=r, column=c, value=val)
            if c == 4:  # Amount column
                cell.number_format = '$#,##0.00'

    ws_archive.column_dimensions['A'].width = 14
    ws_archive.column_dimensions['B'].width = 16
    ws_archive.column_dimensions['C'].width = 38
    ws_archive.column_dimensions['D'].width = 14
    ws_archive.column_dimensions['E'].width = 14

    # --- Sheet 3: Summary ---
    ws_summary = wb.create_sheet('Summary')

    summary_headers = ['Department', 'Total Projects', 'Total Spend', 'Avg Cost']
    for col, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')

    summary_data = [
        ['Operations', 3, 34500.00, 11500.00],
        ['Marketing', 3, 62250.00, 20750.00],
        ['Engineering', 3, 35500.00, 11833.33],
        ['HR', 3, 10400.00, 3466.67],
        ['Finance', 2, 8300.00, 4150.00],
    ]
    for r, row_data in enumerate(summary_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_summary.cell(row=r, column=c, value=val)
            if c in (3, 4):
                cell.number_format = '$#,##0.00'

    ws_summary.column_dimensions['A'].width = 16
    ws_summary.column_dimensions['B'].width = 16
    ws_summary.column_dimensions['C'].width = 14
    ws_summary.column_dimensions['D'].width = 14

    # --- Sheet 4: Data ---
    ws_data = wb.create_sheet('Data')

    data_headers = ['Record ID', 'Timestamp', 'Source', 'Value', 'Notes']
    for col, h in enumerate(data_headers, 1):
        cell = ws_data.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', size=11, bold=True)

    raw_data = [
        ['REC-001', '2024-01-15 09:30', 'System A', 145.67, 'Initial import'],
        ['REC-002', '2024-01-15 10:15', 'System B', 230.44, 'Verified'],
        ['REC-003', '2024-01-16 08:45', 'System A', 98.12, 'Pending review'],
        ['REC-004', '2024-01-16 14:20', 'System C', 312.89, 'Approved'],
        ['REC-005', '2024-01-17 11:00', 'System B', 67.50, 'Flagged'],
        ['REC-006', '2024-01-18 09:00', 'System A', 189.33, 'Verified'],
        ['REC-007', '2024-01-18 16:30', 'System C', 445.00, 'Approved'],
        ['REC-008', '2024-01-19 10:45', 'System B', 123.78, 'Pending review'],
        ['REC-009', '2024-01-20 08:15', 'System A', 276.55, 'Verified'],
        ['REC-010', '2024-01-20 13:00', 'System C', 88.90, 'Initial import'],
    ]
    for r, row_data in enumerate(raw_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_data.cell(row=r, column=c, value=val)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
