"""
Reward Script: Insert hyperlink in Links!A1 pointing to Archive!A1
Task ID: calc_gg3_011
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Cell A1 on Links sheet displays 'Go to Archive'
  Component 2 (0.30): Hyperlink object exists on Links!A1
  Component 3 (0.25): Hyperlink target references Archive sheet cell A1
  Component 4 (0.15): Font shows underline (hyperlink visual indicator)
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_011'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Links' sheet must exist
    if 'Links' not in wb.sheetnames:
        print("FAIL: 'Links' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Links']
    a1 = ws['A1']

    # Component 1: Cell A1 displays 'Go to Archive' (0.30 points)
    # Initial has 'ARCHIVE LINK PLACEHOLDER', golden has 'Go to Archive'
    try:
        val = a1.value
        if val is not None and str(val).strip() == 'Go to Archive':
            print(f"PASS: Component 1 — A1 value is 'Go to Archive' (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 — expected 'Go to Archive', found: {val!r}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Hyperlink object exists on A1 (0.30 points)
    # Initial has no hyperlink, golden has a hyperlink
    try:
        hl = a1.hyperlink
        if hl is not None:
            print(f"PASS: Component 2 — Hyperlink exists on A1 (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 2 — No hyperlink on A1")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Hyperlink target references Archive sheet cell A1 (0.25 points)
    # Golden target is '#Archive!A1' — an internal document link
    try:
        hl = a1.hyperlink
        if hl is not None:
            target = hl.target or ''
            location = hl.location or ''
            # The hyperlink can be stored as target='#Archive!A1' or location='Archive!A1'
            # Normalize and check for 'Archive' sheet reference and 'A1' cell
            combined = (target + ' ' + location).upper()
            has_archive = 'ARCHIVE' in combined
            has_a1 = 'A1' in combined
            if has_archive and has_a1:
                print(f"PASS: Component 3 — Hyperlink points to Archive!A1 (target={target!r}, location={location!r}) (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 3 — Hyperlink target does not reference Archive!A1 (target={target!r}, location={location!r})")
        else:
            print(f"FAIL: Component 3 — No hyperlink to check target")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Font underline is set (hyperlink visual) (0.15 points)
    # Initial has no underline, golden has underline='single'
    try:
        underline = a1.font.underline
        if underline is not None and underline != 'none':
            print(f"PASS: Component 4 — Font underline is '{underline}' (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — Font underline not set (found: {underline!r})")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
