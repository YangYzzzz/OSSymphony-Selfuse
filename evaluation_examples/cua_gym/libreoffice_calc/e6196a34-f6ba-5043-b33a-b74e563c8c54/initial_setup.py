"""
Initial Setup: CRM Contact Scoring Model
Task ID: calc_wf_026
Domain: libreoffice_calc

Creates a Contacts sheet with 30 contacts and columns for Name, Company,
Size, Industry Match, Engagement, Last Contact Date, Score, Rank, Priority.
Score/Rank/Priority columns are left empty for the agent to fill.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from datetime import date, timedelta
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_026'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"

    # Headers
    headers = ['Name', 'Company', 'Size', 'Industry Match', 'Engagement',
               'Last Contact Date', 'Score', 'Rank', 'Priority']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 30 realistic contacts
    contacts = [
        ('Sarah Chen', 'Nextera Solutions', 4, 'Y', 8, date(2026, 3, 15)),
        ('Marcus Johnson', 'Pinnacle Group', 3, 'N', 6, date(2026, 2, 20)),
        ('Elena Rodriguez', 'CloudBridge Inc', 5, 'Y', 9, date(2026, 3, 28)),
        ('James O\'Brien', 'Vertex Analytics', 2, 'Y', 7, date(2026, 1, 10)),
        ('Aisha Patel', 'Meridian Corp', 4, 'N', 5, date(2026, 3, 1)),
        ('David Kim', 'SilverLine Tech', 5, 'Y', 10, date(2026, 3, 30)),
        ('Rachel Foster', 'Atlas Dynamics', 3, 'Y', 4, date(2025, 12, 15)),
        ('Thomas Wright', 'CoreSync Ltd', 1, 'N', 3, date(2026, 2, 5)),
        ('Priya Sharma', 'QuantumEdge', 5, 'Y', 8, date(2026, 3, 20)),
        ('Carlos Martinez', 'BluePeak Systems', 4, 'Y', 7, date(2026, 1, 25)),
        ('Lisa Chang', 'Horizon Labs', 2, 'N', 9, date(2026, 3, 10)),
        ('Robert Taylor', 'Forge Industries', 3, 'Y', 6, date(2026, 2, 14)),
        ('Fatima Al-Hassan', 'Vanguard Digital', 5, 'Y', 8, date(2026, 3, 25)),
        ('Michael Brown', 'TerraWave Corp', 4, 'N', 7, date(2026, 1, 30)),
        ('Jennifer Lee', 'Stratos Consulting', 3, 'Y', 5, date(2026, 2, 28)),
        ('Andre Williams', 'Cipher Networks', 2, 'Y', 8, date(2026, 3, 5)),
        ('Natasha Volkov', 'NovaStar LLC', 5, 'N', 9, date(2026, 3, 18)),
        ('Kevin Murphy', 'Apex Ventures', 4, 'Y', 6, date(2025, 11, 20)),
        ('Diana Okafor', 'Lumina Partners', 3, 'Y', 7, date(2026, 2, 10)),
        ('Ryan Cooper', 'Zenith Group', 1, 'N', 4, date(2026, 3, 12)),
        ('Maria Santos', 'Catalyst Innovations', 5, 'Y', 10, date(2026, 3, 29)),
        ('William Harris', 'Redstone Capital', 4, 'N', 3, date(2025, 12, 1)),
        ('Sophie Laurent', 'Elevate Tech', 2, 'Y', 6, date(2026, 2, 18)),
        ('Benjamin Nakamura', 'Pacific Bridge Co', 3, 'Y', 5, date(2026, 1, 5)),
        ('Amanda Collins', 'Sterling Solutions', 5, 'Y', 9, date(2026, 3, 22)),
        ('Hassan Ibrahim', 'Matrix Consulting', 4, 'N', 7, date(2026, 2, 25)),
        ('Catherine Wu', 'Orion Enterprises', 2, 'Y', 8, date(2026, 3, 8)),
        ('Patrick O\'Neill', 'Summit Analytics', 3, 'N', 6, date(2026, 1, 15)),
        ('Zara Khan', 'Prism Technologies', 5, 'Y', 7, date(2026, 3, 26)),
        ('Daniel Fischer', 'Ironwood Partners', 4, 'Y', 5, date(2026, 2, 8)),
    ]

    for r, (name, company, size, industry, engagement, last_date) in enumerate(contacts, 2):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=company)
        ws.cell(row=r, column=3, value=size)
        ws.cell(row=r, column=4, value=industry)
        ws.cell(row=r, column=5, value=engagement)
        ws.cell(row=r, column=6, value=last_date)
        ws[f'F{r}'].number_format = 'yyyy-mm-dd'
        # G (Score), H (Rank), I (Priority) left empty intentionally

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
