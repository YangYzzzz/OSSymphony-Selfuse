"""
Reward Script: CRM Contact Scoring Model
Task ID: calc_wf_026
Domain: libreoffice_calc
Scoring:
  Component 1: Score formulas in G column (0.25 pts)
  Component 2: Rank formulas in H column (0.15 pts)
  Component 3: Priority formulas in I column (0.10 pts)
  Component 4: 3-color scale conditional formatting on G2:G31 (0.15 pts)
  Component 5: AutoFilter with Priority='High' filter (0.15 pts)
  Component 6: Scatter chart present with score vs days data (0.20 pts)
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_026'


def persist_app_state(domain: str):
    """Try to save any unsaved changes in LibreOffice."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet exists
    if 'Contacts' not in wb.sheetnames:
        print("CRITICAL: 'Contacts' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Contacts']

    # Component 1: Score formulas in G column (0.25 points)
    # The golden file has weighted scoring formulas in G2:G31
    # Formula pattern: =C{r}*2+IF(D{r}="Y",3,0)+E{r}*1.5+MAX(0,10-DAYS(TODAY(),F{r})/30)
    # We check that G cells contain formulas with key elements: weighted factors + IF for industry
    try:
        formula_count = 0
        for r in range(2, 32):
            val = ws.cell(row=r, column=7).value
            if isinstance(val, str) and val.startswith('='):
                # Check for key formula components: multiplication weighting, IF for industry, engagement
                val_upper = val.upper().replace(" ", "")
                has_size_weight = '*2' in val or '2*' in val_upper
                has_if_industry = 'IF(' in val_upper and '"Y"' in val.upper()
                has_engagement = '*1.5' in val or '1.5*' in val_upper
                if has_size_weight and has_if_industry and has_engagement:
                    formula_count += 1
        if formula_count >= 28:  # allow minor tolerance (28 of 30)
            print(f"PASS: Component 1 - Score formulas found in {formula_count}/30 G cells (0.25 pts)")
            total_score += 0.25
        elif formula_count >= 15:
            partial = 0.25 * (formula_count / 30)
            print(f"PARTIAL: Component 1 - Score formulas found in {formula_count}/30 G cells ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - Only {formula_count}/30 G cells have proper score formulas")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Rank formulas in H column (0.15 points)
    # Expected: =RANK(G{r},$G$2:$G$31)
    try:
        rank_count = 0
        for r in range(2, 32):
            val = ws.cell(row=r, column=8).value
            if isinstance(val, str) and val.startswith('='):
                val_upper = val.upper().replace(" ", "")
                if 'RANK(' in val_upper:
                    rank_count += 1
        if rank_count >= 28:
            print(f"PASS: Component 2 - RANK formulas found in {rank_count}/30 H cells (0.15 pts)")
            total_score += 0.15
        elif rank_count >= 15:
            partial = 0.15 * (rank_count / 30)
            print(f"PARTIAL: Component 2 - RANK formulas in {rank_count}/30 H cells ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 - Only {rank_count}/30 H cells have RANK formulas")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Priority formulas in I column (0.10 points)
    # Expected: =IF(H{r}<=10,"High","Low")
    try:
        priority_count = 0
        for r in range(2, 32):
            val = ws.cell(row=r, column=9).value
            if isinstance(val, str) and val.startswith('='):
                val_upper = val.upper().replace(" ", "")
                if 'IF(' in val_upper and 'HIGH' in val.upper():
                    priority_count += 1
        if priority_count >= 28:
            print(f"PASS: Component 3 - Priority formulas found in {priority_count}/30 I cells (0.10 pts)")
            total_score += 0.10
        elif priority_count >= 15:
            partial = 0.10 * (priority_count / 30)
            print(f"PARTIAL: Component 3 - Priority formulas in {priority_count}/30 I cells ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 - Only {priority_count}/30 I cells have Priority formulas")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: 3-color scale conditional formatting on G column (0.15 points)
    # Must have a colorScale rule applied to G2:G31 range with 3 colors
    try:
        color_scale_found = False
        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            # Check if the range covers G column scores
            if 'G' in cf_range:
                for rule in cf.rules:
                    if rule.type == 'colorScale' and rule.colorScale:
                        cs = rule.colorScale
                        if len(cs.cfvo) == 3 and len(cs.color) == 3:
                            color_scale_found = True
                            break
            if color_scale_found:
                break

        if color_scale_found:
            print(f"PASS: Component 4 - 3-color scale conditional formatting on G column (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 - No 3-color scale conditional formatting found on G column")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: AutoFilter with Priority='High' filter (0.15 points)
    # Must have auto_filter defined and a filter on the Priority column for 'High'
    try:
        has_autofilter = False
        has_high_filter = False

        if ws.auto_filter.ref:
            has_autofilter = True
            # Check for filterColumn entries
            if ws.auto_filter.filterColumn:
                for fc in ws.auto_filter.filterColumn:
                    # The Priority column (I) is index 8 (0-based from A)
                    if fc.filters:
                        filter_vals = [f for f in fc.filters.filter]
                        if 'High' in filter_vals:
                            has_high_filter = True

        if has_autofilter and has_high_filter:
            print(f"PASS: Component 5 - AutoFilter with Priority='High' filter applied (0.15 pts)")
            total_score += 0.15
        elif has_autofilter:
            print(f"PARTIAL: Component 5 - AutoFilter defined but no 'High' filter (0.07 pts)")
            total_score += 0.07
        else:
            print(f"FAIL: Component 5 - No AutoFilter defined")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    # Component 6: Scatter chart of score vs days since contact (0.20 points)
    # Must have at least one ScatterChart on the sheet
    try:
        scatter_found = False
        chart_count = len(ws._charts)

        if chart_count > 0:
            for chart in ws._charts:
                chart_type_name = type(chart).__name__
                if chart_type_name == 'ScatterChart':
                    scatter_found = True
                    break

        if scatter_found:
            print(f"PASS: Component 6 - Scatter chart found ({chart_count} chart(s) total) (0.20 pts)")
            total_score += 0.20
        elif chart_count > 0:
            # Some chart exists but not scatter - partial credit
            print(f"PARTIAL: Component 6 - Chart exists but not a scatter chart (type: {type(ws._charts[0]).__name__}) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 6 - No charts found on the sheet")
    except Exception as e:
        print(f"ERROR: Component 6 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
