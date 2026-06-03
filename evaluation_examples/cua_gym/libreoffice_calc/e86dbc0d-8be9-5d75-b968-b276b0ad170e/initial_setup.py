"""
Initial Setup: Create a spreadsheet with 50 rows of sales data for conditional formatting task.
Task ID: calc_gg2_001
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg2_001'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sales'

    # Headers
    headers = ['Order ID', 'Product', 'Region', 'Revenue', 'Cost', 'Profit Margin']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)

    # Product catalog
    products = [
        'Wireless Headphones', 'USB-C Hub', 'Mechanical Keyboard', 'Laptop Stand',
        'Monitor Arm', 'Webcam Pro', 'Desk Lamp', 'External SSD 1TB',
        'Ergonomic Mouse', 'Cable Management Kit', 'Portable Charger',
        'Bluetooth Speaker', 'Screen Protector', 'Phone Mount', 'HDMI Adapter',
        'Noise Cancelling Earbuds', 'Tablet Stylus', 'USB Flash Drive 64GB',
        'Wireless Charging Pad', 'Smart Power Strip'
    ]

    regions = ['North America', 'Europe', 'Asia Pacific', 'Latin America', 'Middle East']

    # Generate 50 rows of realistic data with varied profit margins
    # Ensure a good mix: some <0.10, some 0.10-0.20, some >0.20
    margin_ranges = [
        (0.02, 0.09),   # low margin (red)
        (0.10, 0.20),   # mid margin (yellow)
        (0.21, 0.38),   # high margin (green)
    ]

    for i in range(50):
        row = i + 2
        order_id = f'ORD-2025-{1001 + i}'
        product = products[i % len(products)]
        region = regions[i % len(regions)]

        # Distribute margins: roughly 15 low, 20 mid, 15 high
        if i < 15:
            margin_low, margin_high = margin_ranges[0]
        elif i < 35:
            margin_low, margin_high = margin_ranges[1]
        else:
            margin_low, margin_high = margin_ranges[2]

        margin = round(random.uniform(margin_low, margin_high), 4)
        revenue = round(random.uniform(500, 15000), 2)
        cost = round(revenue * (1 - margin), 2)

        ws.cell(row=row, column=1, value=order_id)
        ws.cell(row=row, column=2, value=product)
        ws.cell(row=row, column=3, value=region)
        ws.cell(row=row, column=4, value=revenue)
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        ws.cell(row=row, column=5, value=cost)
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        ws.cell(row=row, column=6, value=margin)
        ws.cell(row=row, column=6).number_format = '0.00%'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 16

    # NO conditional formatting in initial state
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
