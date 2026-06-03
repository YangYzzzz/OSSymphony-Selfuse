"""
Initial Setup: test_scores.xlsx with student scores (some cells empty)
Task ID: osworld_multi_apps_calc_vscode_010
Domain: libreoffice_calc + vscode
"""

import os
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
DESKTOP = '/home/user/Desktop'
TASK_ID = 'osworld_multi_apps_calc_vscode_010'
OUTPUT = f'{DESKTOP}/test_scores.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    # Ensure Desktop exists
    os.makedirs(DESKTOP, exist_ok=True)

    wb = openpyxl.Workbook()

    # --- Sheet: Scores ---
    ws = wb.active
    ws.title = 'Scores'

    # Header row styling
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    headers = ['StudentName', 'Math', 'Science', 'English', 'History', 'Art']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Student data — some scores are intentionally left empty (None)
    # Empty cells will be filled with subject-level mean by the agent's script
    data = [
        ('Alice Nguyen',     92,   None, 88,   75,   90  ),
        ('Ben Carter',       78,   85,   None, 88,   72  ),
        ('Chloe Martin',     None, 90,   94,   82,   85  ),
        ('David Kim',        88,   76,   72,   None, 68  ),
        ('Elena Foster',     95,   92,   91,   89,   None),
        ('Frank Lopez',      None, 68,   70,   65,   71  ),
        ('Grace Chen',       83,   88,   None, 91,   87  ),
        ('Henry Patel',      70,   None, 75,   78,   80  ),
        ('Isabella Torres',  91,   84,   87,   None, 93  ),
        ('James Wright',     None, 77,   80,   72,   76  ),
        ('Karen Liu',        86,   91,   89,   85,   None),
        ('Liam Scott',       74,   82,   None, 79,   83  ),
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Column widths for readability
    ws.column_dimensions['A'].width = 20
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col_letter].width = 12

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open LibreOffice Calc with test_scores.xlsx
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
