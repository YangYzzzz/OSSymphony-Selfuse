"""
Reward Script: Add total and average rows + bar chart for student exam scores
Task ID: osworld_calc_multi_chart_computed_009
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Total row (row 12) with SUM formulas for 4 subjects     — 0.30 points
  Component 2: Average row (row 13) with AVERAGE formulas for 4 subjects — 0.30 points
  Component 3: Bar chart exists and references the Average row data      — 0.40 points
  Total: 1.0

Initial state: Sheet 'ExamScores', rows 1-11 (header + 10 students), 0 charts.
Golden state: + Total row (row 12) + Average row (row 13) + BarChart using Average row.
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_multi_chart_computed_009'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook — critical precondition gate
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Locate the ExamScores sheet (precondition gate — not scored)
    try:
        if 'ExamScores' in wb.sheetnames:
            ws = wb['ExamScores']
        else:
            ws = wb.active
        print(f"INFO: Using sheet '{ws.title}', max_row={ws.max_row}, max_col={ws.max_column}")
    except Exception as e:
        print(f"CRITICAL: Cannot access sheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # ----------------------------------------------------------------
    # Component 1: Total row (label + SUM formulas for B-E) — 0.30 pts
    # The task requires a "Total" row using SUM formulas appended below
    # the 10 student data rows (i.e., row 12 given the header in row 1).
    # We check:
    #   (a) A12 contains a label that includes 'total' (case-insensitive)
    #   (b) B12, C12, D12, E12 each contain a SUM formula that covers
    #       the data range (references rows 2:11 or equivalent)
    # ----------------------------------------------------------------
    try:
        # Find the Total row: could be row 12 (standard) or nearby
        total_row = None
        for r in range(12, 16):
            label = ws.cell(row=r, column=1).value
            if label and str(label).strip().lower() == 'total':
                total_row = r
                break

        if total_row is None:
            print("FAIL: Component 1 — No 'Total' row label found in A12..A15")
        else:
            a_label = str(ws.cell(row=total_row, column=1).value).strip()
            # Check SUM formulas in columns B-E
            sum_cols_ok = 0
            for col in range(2, 6):  # B=2, C=3, D=4, E=5
                cell_val = ws.cell(row=total_row, column=col).value
                if cell_val and isinstance(cell_val, str):
                    # Accept any SUM formula referencing the data rows
                    if re.search(r'SUM\s*\(', cell_val, re.IGNORECASE):
                        sum_cols_ok += 1
                        print(f"  PASS: SUM formula in col {col} row {total_row}: {cell_val}")
                    else:
                        print(f"  FAIL: Expected SUM formula in col {col} row {total_row}, found: {repr(cell_val)}")
                else:
                    print(f"  FAIL: Expected SUM formula in col {col} row {total_row}, found: {repr(cell_val)}")

            if sum_cols_ok == 4:
                print(f"PASS: Component 1 — Total row at row {total_row} with 4 SUM formulas (0.30 pts)")
                total_score += 0.30
            elif sum_cols_ok >= 2:
                partial = round(0.30 * sum_cols_ok / 4, 2)
                print(f"PARTIAL: Component 1 — Total row has {sum_cols_ok}/4 SUM formulas ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 1 — Total row found at row {total_row} but only {sum_cols_ok}/4 SUM formulas")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ----------------------------------------------------------------
    # Component 2: Average row (label + AVERAGE formulas for B-E) — 0.30 pts
    # The task requires an "Average" row using AVERAGE formulas appended
    # below the Total row (typically row 13).
    # We check:
    #   (a) There is a row with a label containing 'average'
    #   (b) B-E for that row each contain an AVERAGE formula
    # ----------------------------------------------------------------
    try:
        avg_row = None
        for r in range(12, 17):
            label = ws.cell(row=r, column=1).value
            if label and str(label).strip().lower() == 'average':
                avg_row = r
                break

        if avg_row is None:
            print("FAIL: Component 2 — No 'Average' row label found in A12..A16")
        else:
            avg_cols_ok = 0
            for col in range(2, 6):  # B=2, C=3, D=4, E=5
                cell_val = ws.cell(row=avg_row, column=col).value
                if cell_val and isinstance(cell_val, str):
                    if re.search(r'AVERAGE\s*\(', cell_val, re.IGNORECASE):
                        avg_cols_ok += 1
                        print(f"  PASS: AVERAGE formula in col {col} row {avg_row}: {cell_val}")
                    else:
                        print(f"  FAIL: Expected AVERAGE formula in col {col} row {avg_row}, found: {repr(cell_val)}")
                else:
                    print(f"  FAIL: Expected AVERAGE formula in col {col} row {avg_row}, found: {repr(cell_val)}")

            if avg_cols_ok == 4:
                print(f"PASS: Component 2 — Average row at row {avg_row} with 4 AVERAGE formulas (0.30 pts)")
                total_score += 0.30
            elif avg_cols_ok >= 2:
                partial = round(0.30 * avg_cols_ok / 4, 2)
                print(f"PARTIAL: Component 2 — Average row has {avg_cols_ok}/4 AVERAGE formulas ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 2 — Average row found at row {avg_row} but only {avg_cols_ok}/4 AVERAGE formulas")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ----------------------------------------------------------------
    # Component 3: Bar chart exists and uses the Average row data — 0.40 pts
    # The task requires a bar chart comparing average scores by subject.
    # Sub-scoring:
    #   (a) At least 1 chart present in the sheet          — 0.20 pts
    #   (b) Chart type is BarChart (bar or column)          — 0.10 pts
    #   (c) Chart series reference the Average row (row 13
    #       or whichever avg_row was found above)           — 0.10 pts
    # ----------------------------------------------------------------
    try:
        charts = ws._charts
        if not charts:
            print("FAIL: Component 3 — No charts found in the sheet")
        else:
            print(f"INFO: Found {len(charts)} chart(s)")
            chart_score = 0.0

            # Sub-component 3a: chart exists
            chart_score += 0.20
            print("PASS: Component 3a — At least one chart exists (0.20 pts)")

            # Sub-component 3b: is it a BarChart?
            bar_chart_found = False
            for ch in charts:
                if type(ch).__name__ == 'BarChart':
                    bar_chart_found = True
                    print(f"PASS: Component 3b — BarChart found (type='{ch.type}') (0.10 pts)")
                    chart_score += 0.10
                    break
            if not bar_chart_found:
                chart_names = [type(ch).__name__ for ch in charts]
                print(f"FAIL: Component 3b — No BarChart found; chart types: {chart_names}")

            # Sub-component 3c: chart series reference the Average row data
            # The Average row is typically row 13. Series val refs should contain
            # a cell reference in that row (e.g., $B$13, $C$13, etc.)
            avg_row_for_check = avg_row if 'avg_row' in dir() and avg_row else 13
            avg_row_ref_found = False
            for ch in charts:
                for ser in ch.series:
                    try:
                        val_ref = str(ser.val)
                        # Check if the series ref mentions the average row
                        if f'${avg_row_for_check}' in val_ref or f'$B$13' in val_ref or \
                                f'${avg_row_for_check}:' in val_ref or \
                                re.search(rf'\${avg_row_for_check}\b', val_ref):
                            avg_row_ref_found = True
                            print(f"PASS: Component 3c — Series references average row {avg_row_for_check}: {val_ref[:80]}")
                            break
                        # Also accept a range reference in row 13 area
                        if re.search(r'\$[B-E]\$1[23]', val_ref):
                            avg_row_ref_found = True
                            print(f"PASS: Component 3c — Series references avg row area: {val_ref[:80]}")
                            break
                    except Exception:
                        pass
                if avg_row_ref_found:
                    break

            if avg_row_ref_found:
                chart_score += 0.10
                print("PASS: Component 3c — Chart series reference Average row data (0.10 pts)")
            else:
                print(f"FAIL: Component 3c — Chart series do not reference Average row {avg_row_for_check}")

            total_score += chart_score
            print(f"Component 3 total: {chart_score}/0.40 pts")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in a given env
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
