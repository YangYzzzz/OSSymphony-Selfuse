"""
Initial Setup: Student exam score table - pre-task state
Task ID: osworld_calc_multi_chart_computed_009
Domain: libreoffice_calc

Creates a spreadsheet with student exam scores for 10 students across 4 subjects.
No Total/Average rows, no charts (those are the task goals).
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_multi_chart_computed_009'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ExamScores"

    # --- Headers (Row 1) ---
    headers = ['Student', 'Math', 'Science', 'English', 'History']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Student exam data (Rows 2-11) - realistic names and scores ---
    student_data = [
        ['Emily Carter',    88, 92, 85, 79],
        ['James Rodriguez', 75, 68, 90, 83],
        ['Sophia Nguyen',   95, 88, 72, 91],
        ['Liam Thompson',   62, 74, 81, 70],
        ['Olivia Martinez', 84, 91, 88, 77],
        ['Noah Kim',        71, 65, 76, 88],
        ['Ava Patel',       90, 84, 95, 82],
        ['William Zhang',   58, 72, 69, 75],
        ['Isabella Brown',  83, 79, 87, 94],
        ['Ethan Davis',     77, 88, 73, 68],
    ]

    for row_idx, row_data in enumerate(student_data, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="center")

    # --- Column widths ---
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10

    # --- Row height for header ---
    ws.row_dimensions[1].height = 25

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
