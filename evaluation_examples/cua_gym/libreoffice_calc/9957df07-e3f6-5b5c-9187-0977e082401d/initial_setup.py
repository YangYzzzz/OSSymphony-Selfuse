"""
Initial Setup: Price list spreadsheet without data validation on Price column
Task ID: calc_dop_validate_decimal_022
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

WORKDIR = '/home/user'
TASK_ID = 'calc_dop_validate_decimal_022'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: PriceList ---
    ws = wb.active
    ws.title = 'PriceList'

    # Headers in row 1
    headers = ['Item ID', 'Description', 'Supplier', 'Price', 'Effective Date']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
        ws.cell(row=1, column=col).font = Font(bold=True)

    # Realistic price list data — rows 2-200
    # Mix of normal prices, some zero (not yet priced), some negative (data errors),
    # and some exceeding 9999.99
    data_rows = [
        ('ITM-001', 'Wireless Bluetooth Headphones', 'TechSource Inc.', 49.99, date(2025, 1, 15)),
        ('ITM-002', 'USB-C Charging Cable 2m', 'CablePro Ltd.', 12.50, date(2025, 1, 15)),
        ('ITM-003', 'Ergonomic Office Chair', 'FurniMax Corp.', 349.00, date(2025, 2, 1)),
        ('ITM-004', 'Mechanical Keyboard TKL', 'KeyCraft', 89.95, date(2025, 1, 20)),
        ('ITM-005', 'Monitor 27" 4K IPS', 'DisplayTech', 429.99, date(2025, 2, 10)),
        ('ITM-006', 'Portable SSD 1TB', 'StoragePro', 0, date(2025, 1, 15)),           # not yet priced
        ('ITM-007', 'Webcam HD 1080p', 'VisionCam', 75.00, date(2025, 1, 22)),
        ('ITM-008', 'Laptop Stand Aluminum', 'DeskWorks', 38.50, date(2025, 2, 5)),
        ('ITM-009', 'Gaming Mouse RGB', 'PlayGear', -5.00, date(2025, 1, 18)),          # data error
        ('ITM-010', 'Noise Cancelling Earbuds', 'SoundElite', 129.99, date(2025, 2, 3)),
        ('ITM-011', 'Smart LED Desk Lamp', 'LightZone', 54.75, date(2025, 1, 25)),
        ('ITM-012', 'HDMI Cable 4K 3m', 'CablePro Ltd.', 18.99, date(2025, 1, 15)),
        ('ITM-013', 'USB Hub 7-Port', 'ConnectAll', 29.50, date(2025, 2, 8)),
        ('ITM-014', 'Wrist Rest Keyboard Pad', 'ComfortDesk', 22.00, date(2025, 1, 30)),
        ('ITM-015', 'Webcam Privacy Cover', 'SecureView', 4.99, date(2025, 1, 15)),
        ('ITM-016', 'External GPU Dock', 'PowerStation', 10500.00, date(2025, 2, 15)), # exceeds 9999.99
        ('ITM-017', 'Screen Cleaning Kit', 'CleanTech', 11.25, date(2025, 1, 20)),
        ('ITM-018', 'Desk Organizer Set', 'OrganizePro', 34.99, date(2025, 2, 1)),
        ('ITM-019', 'Portable Power Bank 20000mAh', 'PowerBoost', 59.99, date(2025, 1, 28)),
        ('ITM-020', 'Surge Protector 8-Outlet', 'SafeElec', 0, date(2025, 2, 5)),      # not yet priced
        ('ITM-021', 'Vertical Mouse Ergonomic', 'ErgoTech', 49.95, date(2025, 1, 15)),
        ('ITM-022', 'Cable Management Sleeve', 'NeatDesk', 14.50, date(2025, 1, 22)),
        ('ITM-023', 'Trackball Mouse Wireless', 'TrackMaster', 69.00, date(2025, 2, 10)),
        ('ITM-024', 'Laptop Cooling Pad', 'CoolBase', 27.99, date(2025, 1, 18)),
        ('ITM-025', 'VESA Monitor Mount Arm', 'MountPro', 85.00, date(2025, 2, 3)),
        ('ITM-026', 'Keyboard Wrist Support', 'ComfortDesk', 19.99, date(2025, 1, 25)),
        ('ITM-027', 'Mini DisplayPort Adapter', 'CablePro Ltd.', 16.75, date(2025, 1, 15)),
        ('ITM-028', 'Smart Home Hub Bridge', 'HomeConnect', -12.00, date(2025, 2, 8)), # data error
        ('ITM-029', 'Desk Whiteboard 24x36', 'WriteSpace', 45.00, date(2025, 1, 30)),
        ('ITM-030', 'Anti-Glare Screen Filter', 'ClearView', 33.50, date(2025, 1, 20)),
        ('ITM-031', 'USB-A to USB-C Adapter', 'CablePro Ltd.', 8.99, date(2025, 2, 1)),
        ('ITM-032', 'Printer Ink Cartridge Black', 'InkMaster', 24.99, date(2025, 1, 15)),
        ('ITM-033', 'Wireless Charging Pad', 'ChargeFast', 39.99, date(2025, 2, 5)),
        ('ITM-034', 'Security Cable Lock', 'SecureIT', 28.00, date(2025, 1, 22)),
        ('ITM-035', 'Document Scanner Portable', 'ScanPro', 195.00, date(2025, 2, 12)),
        ('ITM-036', 'Noise Machine White Noise', 'SleepWell', 42.50, date(2025, 1, 28)),
        ('ITM-037', 'Mesh Desk Organizer', 'DeskOrder', 21.99, date(2025, 1, 15)),
        ('ITM-038', 'Laptop Docking Station', 'DockMaster', 275.00, date(2025, 2, 3)),
        ('ITM-039', 'Blue Light Glasses', 'VisionGuard', 36.75, date(2025, 1, 20)),
        ('ITM-040', 'Thermal Printer Labels', 'PrintSupply', 0, date(2025, 2, 8)),     # not yet priced
        ('ITM-041', 'Fingerprint Scanner USB', 'SecureScan', 67.00, date(2025, 1, 25)),
        ('ITM-042', 'Conference Speakerphone', 'MeetTech', 159.99, date(2025, 1, 15)),
        ('ITM-043', 'Rolling File Cabinet', 'OfficeMax', 189.50, date(2025, 2, 1)),
        ('ITM-044', 'Letter Tray Stackable Set', 'DeskOrder', 17.25, date(2025, 1, 18)),
        ('ITM-045', 'Desktop UPS Battery Backup', 'PowerGuard', 119.99, date(2025, 2, 10)),
        ('ITM-046', 'Mechanical Pencil Professional', 'WriteRight', 6.50, date(2025, 1, 22)),
        ('ITM-047', 'Barcode Scanner Wireless', 'ScanMaster', 145.00, date(2025, 1, 30)),
        ('ITM-048', 'Projector Screen 100"', 'ScreenView', 11200.00, date(2025, 2, 5)),# exceeds 9999.99
        ('ITM-049', 'Smart Plug Wi-Fi Enabled', 'HomeAuto', 22.99, date(2025, 1, 15)),
        ('ITM-050', 'Surge Protector Power Strip', 'SafeElec', 35.00, date(2025, 2, 3)),
        ('ITM-051', 'Pen Holder Magnetic', 'DeskOrder', 13.50, date(2025, 1, 20)),
        ('ITM-052', 'Ergonomic Footrest', 'ComfortDesk', 49.00, date(2025, 2, 8)),
        ('ITM-053', 'Calculator Scientific', 'MathPro', 28.75, date(2025, 1, 25)),
        ('ITM-054', 'Adjustable Monitor Riser', 'DeskWorks', 52.00, date(2025, 1, 15)),
        ('ITM-055', 'Paper Shredder Cross-Cut', 'DocDestroy', 89.99, date(2025, 2, 1)),
        ('ITM-056', 'Label Maker Handheld', 'LabelPro', 44.50, date(2025, 1, 22)),
        ('ITM-057', 'Wireless Presentation Clicker', 'PresentPro', 31.99, date(2025, 2, 12)),
        ('ITM-058', 'Anti-Fatigue Floor Mat', 'StandSafe', 67.50, date(2025, 1, 28)),
        ('ITM-059', 'Cable Tray Under Desk', 'NeatDesk', 38.00, date(2025, 1, 18)),
        ('ITM-060', 'Bookend Set Heavy Duty', 'ShelfMaster', 0, date(2025, 2, 5)),     # not yet priced
        ('ITM-061', 'Desk Drawer Organizer', 'OrganizePro', 16.99, date(2025, 1, 15)),
        ('ITM-062', 'Monitor Calibration Tool', 'ColorTrue', 225.00, date(2025, 2, 10)),
        ('ITM-063', 'Ethernet Cable Cat6 10m', 'NetCable', 19.50, date(2025, 1, 20)),
        ('ITM-064', 'Screen Privacy Filter 24"', 'ClearView', 58.75, date(2025, 1, 30)),
        ('ITM-065', 'Hand Sanitizer Dispenser', 'CleanZone', -2.50, date(2025, 2, 3)), # data error
        ('ITM-066', 'Keyboard Cover Silicone', 'KeyGuard', 9.99, date(2025, 1, 25)),
        ('ITM-067', 'Standing Desk Converter', 'StandUp', 349.99, date(2025, 1, 15)),
        ('ITM-068', 'Business Card Scanner', 'ScanPro', 89.00, date(2025, 2, 8)),
        ('ITM-069', 'Lumbar Support Cushion', 'ComfortDesk', 42.00, date(2025, 1, 22)),
        ('ITM-070', 'VGA to HDMI Converter', 'CablePro Ltd.', 21.50, date(2025, 2, 1)),
        ('ITM-071', 'Desk Calendar Planner 2025', 'TimePlan', 14.99, date(2025, 1, 15)),
        ('ITM-072', 'Foldable Laptop Stand', 'DeskWorks', 33.00, date(2025, 1, 18)),
        ('ITM-073', 'Sticky Note Dispenser', 'DeskOrder', 7.25, date(2025, 2, 5)),
        ('ITM-074', 'Server Rack 12U Wall Mount', 'RackTech', 13500.00, date(2025, 1, 28)), # exceeds
        ('ITM-075', 'Eye Drops Computer Vision', 'VisionCare', 0, date(2025, 2, 10)),   # not priced
        ('ITM-076', 'Ambient Light Bar Monitor', 'LightZone', 29.99, date(2025, 1, 20)),
        ('ITM-077', 'Laptop Privacy Screen 15"', 'SecureView', 47.50, date(2025, 1, 25)),
        ('ITM-078', 'Multi-Device Bluetooth Keyboard', 'KeyCraft', 99.00, date(2025, 2, 3)),
        ('ITM-079', 'Touchpad Numeric Keypad', 'KeyCraft', 55.75, date(2025, 1, 15)),
        ('ITM-080', 'Document Holder Clip-On', 'DeskOrder', 18.00, date(2025, 2, 8)),
        ('ITM-081', 'Air Purifier Desk Size', 'CleanAir', 79.99, date(2025, 1, 22)),
        ('ITM-082', 'USB-C Hub 9-in-1', 'ConnectAll', 65.00, date(2025, 2, 1)),
        ('ITM-083', 'Mouse Bungee Holder', 'PlayGear', 15.50, date(2025, 1, 18)),
        ('ITM-084', 'Headset Stand Organizer', 'AudioDesk', 23.99, date(2025, 2, 12)),
        ('ITM-085', 'Cable Label Kit', 'NeatDesk', 11.00, date(2025, 1, 28)),
        ('ITM-086', 'Fax Machine All-in-One', 'OfficePrint', 199.50, date(2025, 1, 15)),
        ('ITM-087', 'Telephone Headset Wired', 'AudioDesk', 56.25, date(2025, 2, 5)),
        ('ITM-088', 'Stationery Set Premium', 'WriteRight', 32.00, date(2025, 1, 20)),
        ('ITM-089', 'Laptop Bag 15.6"', 'CarryPro', 58.99, date(2025, 1, 30)),
        ('ITM-090', 'Monitor Backlight LED', 'LightZone', 0, date(2025, 2, 3)),         # not priced
        ('ITM-091', 'Whiteboard Markers Set', 'WriteSpace', 12.75, date(2025, 1, 25)),
        ('ITM-092', 'Name Badge Holder Set 50', 'BadgePro', 19.00, date(2025, 1, 15)),
        ('ITM-093', 'Desk Nameplate Holder', 'OrganizePro', 10.50, date(2025, 2, 8)),
        ('ITM-094', 'Portable Bluetooth Speaker', 'SoundElite', 88.99, date(2025, 1, 22)),
        ('ITM-095', 'Noise Isolation Headphones', 'SoundElite', 219.00, date(2025, 2, 1)),
        ('ITM-096', 'Video Conference Camera 4K', 'VisionCam', 399.99, date(2025, 1, 18)),
        ('ITM-097', 'Smart Whiteboard 65"', 'WriteSpace', 12800.00, date(2025, 2, 15)),# exceeds
        ('ITM-098', 'Binding Machine Comb', 'DocFinish', 78.50, date(2025, 1, 25)),
        ('ITM-099', 'Telephone Base Unit', 'CallPro', 95.00, date(2025, 1, 15)),
        ('ITM-100', 'Desktop Stapler Heavy Duty', 'OfficeMax', 36.75, date(2025, 2, 10)),
        ('ITM-101', 'Laptop Backpack 17"', 'CarryPro', 72.00, date(2025, 1, 20)),
        ('ITM-102', 'Mouse Pad XL Extended', 'PlayGear', 25.99, date(2025, 1, 30)),
        ('ITM-103', 'Computer Cart Mobile', 'FurniMax Corp.', -8.00, date(2025, 2, 5)),# data error
        ('ITM-104', 'USB-A 3.0 Flash Drive 64GB', 'StoragePro', 18.50, date(2025, 1, 22)),
        ('ITM-105', 'Laminator A4 Hot', 'DocFinish', 49.99, date(2025, 2, 3)),
        ('ITM-106', 'Presentation Pointer Laser', 'PresentPro', 27.50, date(2025, 1, 15)),
        ('ITM-107', 'Folding Table 6-Foot', 'FurniMax Corp.', 89.00, date(2025, 2, 8)),
        ('ITM-108', 'Staple Remover Claw Type', 'OfficeMax', 5.25, date(2025, 1, 25)),
        ('ITM-109', 'Paper Tray A4 Capacity', 'DeskOrder', 22.50, date(2025, 1, 18)),
        ('ITM-110', 'Ink Roller Stamp Pad Blue', 'InkMaster', 0, date(2025, 2, 1)),     # not priced
        ('ITM-111', 'Standing Desk Electric', 'StandUp', 799.99, date(2025, 1, 15)),
        ('ITM-112', 'Drawer Pedestal 3-Drawer', 'FurniMax Corp.', 245.00, date(2025, 2, 12)),
        ('ITM-113', 'Glare-Free LED Ceiling Light', 'LightZone', 145.50, date(2025, 1, 28)),
        ('ITM-114', 'Bookcase 5-Shelf White', 'ShelfMaster', 179.99, date(2025, 1, 20)),
        ('ITM-115', 'Office Partition Panel', 'SpaceDiv', 425.00, date(2025, 2, 5)),
        ('ITM-116', 'Rolling Whiteboard Stand', 'WriteSpace', 275.50, date(2025, 1, 22)),
        ('ITM-117', 'Chair Mat Hardwood Floor', 'FloorPro', 59.00, date(2025, 2, 3)),
        ('ITM-118', 'Fingerprint Time Clock', 'AttendTrack', 189.99, date(2025, 1, 15)),
        ('ITM-119', 'Receipt Printer Thermal', 'PrintSupply', 155.00, date(2025, 1, 25)),
        ('ITM-120', 'Cash Drawer Printer Connect', 'RetailPro', 0, date(2025, 2, 8)),   # not priced
        ('ITM-121', 'Conference Table 8-Person', 'FurniMax Corp.', 15999.00, date(2025, 1, 18)), # exceeds
        ('ITM-122', 'Network Switch 24-Port', 'NetGear', 299.99, date(2025, 2, 1)),
        ('ITM-123', 'Keyboard Tray Under Desk', 'DeskWorks', 68.50, date(2025, 1, 30)),
        ('ITM-124', 'Monitor Extension Arm Dual', 'MountPro', 135.00, date(2025, 1, 15)),
        ('ITM-125', 'Wireless Keyboard Compact', 'KeyCraft', 45.99, date(2025, 2, 10)),
        ('ITM-126', 'Pen Drive USB-C 128GB', 'StoragePro', 29.00, date(2025, 1, 20)),
        ('ITM-127', 'Highlighter Set 10 Colors', 'WriteRight', 14.50, date(2025, 1, 25)),
        ('ITM-128', 'Desktop Bookstand Adjustable', 'ShelfMaster', 31.75, date(2025, 2, 3)),
        ('ITM-129', 'LCD Writing Tablet 12"', 'DigiWrite', -3.00, date(2025, 1, 22)),   # data error
        ('ITM-130', 'Clipboard A4 Hard Cover', 'WriteRight', 8.99, date(2025, 2, 8)),
        ('ITM-131', 'Stamp Self-Inking PAID', 'InkMaster', 12.00, date(2025, 1, 15)),
        ('ITM-132', 'Portfolio Binder A4 Zip', 'CarryPro', 24.50, date(2025, 1, 28)),
        ('ITM-133', 'Tape Dispenser Desktop', 'OfficeMax', 11.75, date(2025, 2, 5)),
        ('ITM-134', 'Glue Stick Set 12 Pack', 'ArtSupply', 7.50, date(2025, 1, 18)),
        ('ITM-135', 'Index Tab Dividers 5-Set', 'DeskOrder', 5.99, date(2025, 2, 1)),
        ('ITM-136', 'Paper Cutter A3 Guillotine', 'DocFinish', 79.00, date(2025, 1, 22)),
        ('ITM-137', 'Corner Bookcase L-Shaped', 'ShelfMaster', 0, date(2025, 2, 10)),   # not priced
        ('ITM-138', 'Fire Proof Safe Box Small', 'SecureIT', 199.99, date(2025, 1, 15)),
        ('ITM-139', 'Motion Sensor Lamp USB', 'LightZone', 18.50, date(2025, 1, 30)),
        ('ITM-140', 'Cork Notice Board 90x60', 'WriteSpace', 44.00, date(2025, 2, 3)),
        ('ITM-141', 'Drawer Keyboard Slide Out', 'DeskWorks', 55.00, date(2025, 1, 25)),
        ('ITM-142', 'Laptop Sleeve 14" Neoprene', 'CarryPro', 21.99, date(2025, 1, 15)),
        ('ITM-143', 'Pop-Up Sticky Note Cube', 'DeskOrder', 4.25, date(2025, 2, 8)),
        ('ITM-144', 'Cable Protector Sleeve Coil', 'NeatDesk', 6.99, date(2025, 1, 20)),
        ('ITM-145', 'Paperclip Dispenser Magnetic', 'OfficeMax', 9.50, date(2025, 2, 5)),
        ('ITM-146', 'USB-C Power Delivery Adapter 100W', 'PowerBoost', 39.99, date(2025, 1, 22)),
        ('ITM-147', 'Wireless Earpiece Bluetooth', 'AudioDesk', 59.00, date(2025, 1, 28)),
        ('ITM-148', 'Desk Fan USB Powered', 'CoolBase', 17.50, date(2025, 2, 1)),
        ('ITM-149', 'Phone Holder Desk Ring Clamp', 'MountPro', 15.25, date(2025, 1, 15)),
        ('ITM-150', 'Ergonomic Backrest Cushion', 'ComfortDesk', 37.99, date(2025, 2, 12)),
        ('ITM-151', 'Presentation Folder Branded', 'CarryPro', 0, date(2025, 1, 18)),   # not priced
        ('ITM-152', 'Rubber Band Assortment Pack', 'OfficeMax', 3.50, date(2025, 2, 5)),
        ('ITM-153', 'Magnetic Cabinet Lock', 'SecureIT', 14.75, date(2025, 1, 25)),
        ('ITM-154', 'Battery Charger AA/AAA', 'PowerBoost', 22.00, date(2025, 1, 15)),
        ('ITM-155', 'Computer Trolley Cart', 'FurniMax Corp.', 145.50, date(2025, 2, 3)),
        ('ITM-156', 'Padlock Combination Steel', 'SecureIT', 16.99, date(2025, 1, 20)),
        ('ITM-157', 'Folding Chair Stackable', 'FurniMax Corp.', 69.00, date(2025, 1, 30)),
        ('ITM-158', 'Copier Paper A4 500 Sheets', 'PrintSupply', 9.99, date(2025, 2, 8)),
        ('ITM-159', 'Frame 8x10 Certificate', 'WriteSpace', 13.25, date(2025, 1, 22)),
        ('ITM-160', 'Pen Set Ballpoint 12 Pack', 'WriteRight', 8.50, date(2025, 2, 1)),
        ('ITM-161', 'Spiral Notebook A5', 'WriteRight', 6.75, date(2025, 1, 15)),
        ('ITM-162', 'Correction Tape Roller', 'OfficeMax', 3.99, date(2025, 1, 25)),
        ('ITM-163', 'Double-Sided Tape 12mm', 'ArtSupply', 4.50, date(2025, 2, 5)),
        ('ITM-164', 'Permanent Marker Set Black', 'InkMaster', 11.25, date(2025, 1, 18)),
        ('ITM-165', 'Ring Binder 2" White', 'DeskOrder', 5.00, date(2025, 2, 10)),
        ('ITM-166', 'Document Binding Covers Clear', 'DocFinish', 18.00, date(2025, 1, 28)),
        ('ITM-167', 'Filing Box Archival Quality', 'ShelfMaster', 0, date(2025, 1, 15)),# not priced
        ('ITM-168', 'Stencil Set Office Text', 'WriteRight', 7.25, date(2025, 2, 3)),
        ('ITM-169', 'Window Envelope C4 Box 250', 'OfficeMax', 19.50, date(2025, 1, 22)),
        ('ITM-170', 'Postal Scale Digital 5kg', 'WeighPro', 45.00, date(2025, 2, 8)),
        ('ITM-171', 'Mailing Tube 5-Pack', 'OfficeMax', 12.99, date(2025, 1, 20)),
        ('ITM-172', 'Bubble Wrap Roll 30m', 'PackSafe', 28.50, date(2025, 1, 30)),
        ('ITM-173', 'Strapping Tape Heavy Duty', 'PackSafe', 8.75, date(2025, 2, 1)),
        ('ITM-174', 'Box Sealing Tape Clear 6-Pack', 'PackSafe', -1.00, date(2025, 1, 25)),# data error
        ('ITM-175', 'Padded Envelope A4 Pack 10', 'PackSafe', 15.50, date(2025, 1, 15)),
        ('ITM-176', 'Electric Stapler Desktop', 'OfficeMax', 54.99, date(2025, 2, 12)),
        ('ITM-177', 'Ream Sticker Labels A4', 'LabelPro', 16.25, date(2025, 1, 18)),
        ('ITM-178', 'Binding Spiral Wire 14mm', 'DocFinish', 23.00, date(2025, 2, 5)),
        ('ITM-179', 'Sheet Music Stand Adjustable', 'AudioDesk', 0, date(2025, 1, 22)),  # not priced
        ('ITM-180', 'Floor Lamp LED Adjustable', 'LightZone', 105.00, date(2025, 1, 28)),
        ('ITM-181', 'Conference Chair Wheeled', 'FurniMax Corp.', 299.00, date(2025, 1, 15)),
        ('ITM-182', 'Monitor Privacy Hood 27"', 'SecureView', 75.50, date(2025, 2, 3)),
        ('ITM-183', 'Wall Calendar Magnetic', 'TimePlan', 18.75, date(2025, 1, 20)),
        ('ITM-184', 'Mouse Pad with Wireless Charging', 'ChargeFast', 49.99, date(2025, 2, 8)),
        ('ITM-185', 'Ergonomic Keyboard Vertical', 'ErgoTech', 139.00, date(2025, 1, 25)),
        ('ITM-186', 'Gel Wrist Pad Mouse', 'ComfortDesk', 14.99, date(2025, 1, 15)),
        ('ITM-187', 'Filing Cabinet 4-Drawer Steel', 'FurniMax Corp.', 389.00, date(2025, 2, 1)),
        ('ITM-188', 'Photocopier Toner Black', 'InkMaster', 89.50, date(2025, 1, 18)),
        ('ITM-189', 'Inkjet Printer Color', 'PrintSupply', 299.99, date(2025, 2, 10)),
        ('ITM-190', 'Digital Picture Frame 10"', 'DisplayTech', 0, date(2025, 1, 22)),  # not priced
        ('ITM-191', 'Laser Printer Mono A4', 'PrintSupply', 399.00, date(2025, 1, 30)),
        ('ITM-192', 'Scanner Flatbed A4 Color', 'ScanPro', 185.50, date(2025, 2, 5)),
        ('ITM-193', 'Plotter Printer A1', 'PrintSupply', 9999.99, date(2025, 1, 15)),   # at boundary
        ('ITM-194', 'Label Printer Industrial', 'LabelPro', 249.00, date(2025, 2, 3)),
        ('ITM-195', 'Projector DLP 3000 Lumens', 'DisplayTech', 799.99, date(2025, 1, 20)),
        ('ITM-196', 'Interactive Display 86"', 'DisplayTech', 14500.00, date(2025, 1, 28)), # exceeds
        ('ITM-197', 'Whiteboard Interactive Smart', 'WriteSpace', 0, date(2025, 2, 8)),  # not priced
        ('ITM-198', 'Presentation Remote 2.4GHz', 'PresentPro', 25.50, date(2025, 1, 15)),
        ('ITM-199', 'AV Cart Mobile Adjustable', 'FurniMax Corp.', 235.00, date(2025, 2, 1)),
    ]

    for r, row_data in enumerate(data_rows, 2):
        ws.cell(row=r, column=1, value=row_data[0])  # Item ID
        ws.cell(row=r, column=2, value=row_data[1])  # Description
        ws.cell(row=r, column=3, value=row_data[2])  # Supplier
        ws.cell(row=r, column=4, value=row_data[3])  # Price
        ws.cell(row=r, column=5, value=row_data[4])  # Effective Date

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 16

    # NO data validation on column D (that's what the task requires adding)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet: PriceList with {len(data_rows)} data rows (rows 2-200)')


create_initial()
