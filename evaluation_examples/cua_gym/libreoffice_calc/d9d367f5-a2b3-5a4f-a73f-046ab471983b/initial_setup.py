"""
Initial Setup: Add data bars to Budget Overview spreadsheet
Task ID: calc_fmt_condfmt_databar_044
Domain: libreoffice_calc

Creates a 'Budget Overview' sheet with department data.
NO conditional formatting exists initially.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_condfmt_databar_044'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Budget Overview sheet ---
    ws = wb.active
    ws.title = 'Budget Overview'

    # Headers
    headers = ['Department', 'Headcount', 'Budget ($K)', 'Spent ($K)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.alignment = Alignment(horizontal='center')

    # Realistic department data (14 rows, C values range 50-800)
    data = [
        ['Engineering',       142,  800,  763],
        ['Product Management',  28,  320,  298],
        ['Sales',              85,  650,  612],
        ['Marketing',          34,  420,  387],
        ['Human Resources',    22,  180,  145],
        ['Finance',            19,  210,  198],
        ['Legal',              11,  150,  129],
        ['Operations',         56,  380,  356],
        ['Customer Support',   63,  290,  271],
        ['Research',           47,  520,  483],
        ['IT Infrastructure',  38,  340,  312],
        ['Design',             25,  175,  162],
        ['Data Analytics',     31,  260,  241],
        ['Executive',           9,   50,   48],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14

    # NO conditional formatting — this is the initial state before the task

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
