"""
Initial Setup: Apply borders to header and data ranges in a department report
Task ID: calc_ggf_015
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_015'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Report'

    # --- Headers in A1:H1 ---
    headers = [
        'Department', 'Employee', 'Q1 Revenue', 'Q2 Revenue',
        'Q3 Revenue', 'Q4 Revenue', 'Annual Total', 'Region'
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # --- 49 rows of realistic data (A2:H50) ---
    departments = [
        'Engineering', 'Marketing', 'Sales', 'Finance', 'Operations',
        'Human Resources', 'Legal', 'Customer Support', 'Research', 'Product'
    ]
    employees = [
        'Sarah Chen', 'Marcus Johnson', 'Priya Patel', 'James O\'Brien',
        'Yuki Tanaka', 'Elena Rodriguez', 'David Kim', 'Fatima Al-Hassan',
        'Robert Taylor', 'Amara Okafor', 'Lucas Fernandez', 'Mei Lin Wang',
        'Thomas Mueller', 'Aisha Begum', 'Carlos Vega', 'Hannah Schmidt',
        'Omar Farouk', 'Sofia Petrov', 'Benjamin Wright', 'Nadia Kowalski',
        'Daniel Park', 'Isabella Moretti', 'Victor Andersen', 'Grace Nakamura',
        'Ahmed Rashid', 'Catherine Dubois', 'Kevin O\'Sullivan', 'Lena Johansson',
        'Miguel Santos', 'Rachel Goldstein', 'Dmitri Volkov', 'Emma Larsson',
        'Hassan Ali', 'Chloe Martin', 'Jack Thompson', 'Zara Hussein',
        'Patrick Brennan', 'Ananya Sharma', 'William Foster', 'Ingrid Bergman',
        'Samuel Osei', 'Laura Bianchi', 'Nathan Clarke', 'Rosa Gutierrez',
        'Philip Hawkins', 'Julia Novak', 'George Patterson', 'Sana Malik',
        'Ryan Cooper'
    ]
    regions = [
        'North America', 'Europe', 'Asia Pacific', 'Latin America',
        'Middle East', 'Africa'
    ]

    import random
    random.seed(42)

    for i in range(49):
        row = i + 2
        dept = departments[i % len(departments)]
        emp = employees[i]
        q1 = round(random.uniform(15000, 95000), 2)
        q2 = round(random.uniform(15000, 95000), 2)
        q3 = round(random.uniform(15000, 95000), 2)
        q4 = round(random.uniform(15000, 95000), 2)
        total = round(q1 + q2 + q3 + q4, 2)
        region = regions[i % len(regions)]

        ws.cell(row=row, column=1, value=dept)
        ws.cell(row=row, column=2, value=emp)
        ws.cell(row=row, column=3, value=q1)
        ws.cell(row=row, column=4, value=q2)
        ws.cell(row=row, column=5, value=q3)
        ws.cell(row=row, column=6, value=q4)
        ws.cell(row=row, column=7, value=total)
        ws.cell(row=row, column=8, value=region)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 16

    # NO borders applied - this is the initial borderless state
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
