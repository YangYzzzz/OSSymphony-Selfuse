"""
Reward Script: Apply borders to header and data ranges in LibreOffice Calc
Task ID: calc_ggf_015
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20): Header row A1:H1 has thick top border
  Component 2 (0.20): Header row has thick outer left (A1) and right (H1) borders
  Component 3 (0.25): Header row A1:H1 has double bottom border
  Component 4 (0.25): Data range A2:H50 has thin inner borders
  Component 5 (0.10): No borders outside the A1:H50 range
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_015'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Report' sheet must exist
    if 'Report' not in wb.sheetnames:
        print("CRITICAL: 'Report' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Report']

    # Component 1: Header row A1:H1 has thick top border (0.20 points)
    try:
        thick_top_count = 0
        for col in range(1, 9):
            cell = ws.cell(row=1, column=col)
            if cell.border.top.style == 'thick':
                thick_top_count += 1
        if thick_top_count == 8:
            print(f"PASS: Component 1 - All 8 header cells have thick top border (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 - {thick_top_count}/8 header cells have thick top border")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Header row has thick outer left (A1) and thick outer right (H1) (0.20 points)
    try:
        left_ok = ws.cell(row=1, column=1).border.left.style == 'thick'
        right_ok = ws.cell(row=1, column=8).border.right.style == 'thick'
        if left_ok and right_ok:
            print(f"PASS: Component 2 - Thick left on A1 and thick right on H1 (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 - A1 left={ws.cell(row=1, column=1).border.left.style}, H1 right={ws.cell(row=1, column=8).border.right.style}")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Header row A1:H1 has double bottom border (0.25 points)
    try:
        double_bottom_count = 0
        for col in range(1, 9):
            cell = ws.cell(row=1, column=col)
            if cell.border.bottom.style == 'double':
                double_bottom_count += 1
        if double_bottom_count == 8:
            print(f"PASS: Component 3 - All 8 header cells have double bottom border (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 3 - {double_bottom_count}/8 header cells have double bottom border")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Data range A2:H50 has thin inner borders (0.25 points)
    # Sample cells across the data range to verify thin borders
    try:
        sample_rows = [2, 10, 25, 40, 50]
        sample_cols = [1, 3, 5, 8]
        thin_count = 0
        total_checked = 0
        for row in sample_rows:
            for col in sample_cols:
                cell = ws.cell(row=row, column=col)
                b = cell.border
                has_thin = True
                # Check all four sides are thin
                for side_name, side in [('left', b.left), ('right', b.right), ('top', b.top), ('bottom', b.bottom)]:
                    if side.style != 'thin':
                        has_thin = False
                if has_thin:
                    thin_count += 1
                total_checked += 1

        # Require at least 90% of sampled cells to have thin borders
        ratio = thin_count / total_checked if total_checked > 0 else 0
        if ratio >= 0.9:
            print(f"PASS: Component 4 - {thin_count}/{total_checked} sampled data cells have thin borders (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 4 - Only {thin_count}/{total_checked} sampled data cells have thin borders")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Borders are contained within A1:H50 only — no borders outside range
    # AND at least some borders exist inside the range (anchors to task change)
    # (0.10 points)
    try:
        # First check that borders exist inside the range (task was done)
        inside_has_borders = ws.cell(row=1, column=1).border.top.style is not None
        if not inside_has_borders:
            print(f"FAIL: Component 5 - No borders inside range, so containment check is moot")
        else:
            outside_border_found = False
            # Check row 51 cols 1-8
            for col in range(1, 9):
                cell = ws.cell(row=51, column=col)
                b = cell.border
                for side in [b.left, b.right, b.top, b.bottom]:
                    if side.style is not None:
                        outside_border_found = True
                        break
                if outside_border_found:
                    break
            # Check col 9 rows 1-5
            if not outside_border_found:
                for row in range(1, 6):
                    cell = ws.cell(row=row, column=9)
                    b = cell.border
                    for side in [b.left, b.right, b.top, b.bottom]:
                        if side.style is not None:
                            outside_border_found = True
                            break
                    if outside_border_found:
                        break

            if not outside_border_found:
                print(f"PASS: Component 5 - Borders contained within A1:H50 (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 5 - Found borders outside the expected range")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
