"""
Initial Setup: Quiz/Test Generator and Grader - Pre-task state
Task ID: calc_wf_058
Domain: libreoffice_calc

Creates a workbook with:
- "Bank" sheet: 30 multiple-choice questions (visible in initial state)
- "Test" sheet: Layout for 10 questions, answer input area, score section (no formulas yet)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_058'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # =========================================================
    # Sheet 1: Bank — 30 multiple-choice questions
    # =========================================================
    ws_bank = wb.active
    ws_bank.title = 'Bank'

    # Headers
    bank_headers = ['Question', 'Option A', 'Option B', 'Option C', 'Option D', 'Correct Answer']
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(bank_headers, 1):
        cell = ws_bank.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 30 questions — realistic general knowledge / science / math
    questions = [
        ["What is the chemical symbol for gold?", "Au", "Ag", "Fe", "Cu", "A"],
        ["Which planet is closest to the Sun?", "Venus", "Mercury", "Mars", "Earth", "B"],
        ["What is the square root of 144?", "10", "14", "12", "11", "C"],
        ["Who wrote 'Romeo and Juliet'?", "Shakespeare", "Dickens", "Austen", "Hemingway", "A"],
        ["What is the boiling point of water in Celsius?", "90", "110", "80", "100", "D"],
        ["Which gas do plants absorb from the atmosphere?", "Oxygen", "Carbon Dioxide", "Nitrogen", "Hydrogen", "B"],
        ["What is 15% of 200?", "25", "35", "30", "20", "C"],
        ["Which organ pumps blood through the body?", "Heart", "Liver", "Lungs", "Brain", "A"],
        ["What year did World War II end?", "1943", "1944", "1946", "1945", "D"],
        ["What is the largest ocean on Earth?", "Atlantic", "Pacific", "Indian", "Arctic", "B"],
        ["How many sides does a hexagon have?", "5", "7", "8", "6", "D"],
        ["What is the speed of light approximately?", "300,000 km/s", "150,000 km/s", "500,000 km/s", "100,000 km/s", "A"],
        ["Which element has the atomic number 1?", "Helium", "Hydrogen", "Lithium", "Carbon", "B"],
        ["What is the capital of Japan?", "Osaka", "Kyoto", "Tokyo", "Nagoya", "C"],
        ["Who painted the Mona Lisa?", "Leonardo da Vinci", "Michelangelo", "Raphael", "Donatello", "A"],
        ["What is the powerhouse of the cell?", "Nucleus", "Ribosome", "Mitochondria", "Golgi Apparatus", "C"],
        ["How many continents are there?", "5", "6", "8", "7", "D"],
        ["What is the chemical formula for water?", "H2O", "CO2", "NaCl", "O2", "A"],
        ["Which country is known as the Land of the Rising Sun?", "China", "Japan", "Thailand", "Korea", "B"],
        ["What is the derivative of x squared?", "x", "2", "2x", "x/2", "C"],
        ["What is the largest mammal on Earth?", "Blue Whale", "Elephant", "Giraffe", "Hippopotamus", "A"],
        ["How many bones are in the adult human body?", "196", "206", "216", "186", "B"],
        ["What is the currency of the United Kingdom?", "Euro", "Dollar", "Pound Sterling", "Franc", "C"],
        ["Which vitamin is produced when skin is exposed to sunlight?", "Vitamin D", "Vitamin C", "Vitamin A", "Vitamin B12", "A"],
        ["What is the freezing point of water in Fahrenheit?", "0", "32", "100", "212", "B"],
        ["How many degrees are in a circle?", "180", "270", "360", "90", "C"],
        ["What is the hardest natural substance?", "Diamond", "Ruby", "Sapphire", "Emerald", "A"],
        ["Which blood type is the universal donor?", "A", "O negative", "AB", "B", "B"],
        ["What is the SI unit of force?", "Joule", "Watt", "Newton", "Pascal", "C"],
        ["Who developed the theory of relativity?", "Albert Einstein", "Isaac Newton", "Niels Bohr", "Max Planck", "A"],
    ]

    data_font = Font(name='Arial', size=10)
    data_align = Alignment(vertical='center', wrap_text=True)
    center_align = Alignment(horizontal='center', vertical='center')

    for r, q in enumerate(questions, 2):
        for c, val in enumerate(q, 1):
            cell = ws_bank.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = thin_border
            if c == 1:
                cell.alignment = data_align
            elif c == 6:
                cell.alignment = center_align
            else:
                cell.alignment = Alignment(vertical='center')

    # Column widths
    ws_bank.column_dimensions['A'].width = 55
    ws_bank.column_dimensions['B'].width = 22
    ws_bank.column_dimensions['C'].width = 22
    ws_bank.column_dimensions['D'].width = 22
    ws_bank.column_dimensions['E'].width = 22
    ws_bank.column_dimensions['F'].width = 16

    # =========================================================
    # Sheet 2: Test — Layout for quiz (no formulas in initial)
    # =========================================================
    ws_test = wb.create_sheet('Test')

    # Title
    ws_test.merge_cells('A1:H1')
    title_cell = ws_test['A1']
    title_cell.value = 'Quiz / Test Sheet'
    title_cell.font = Font(name='Arial', size=16, bold=True, color='2F5496')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_test.row_dimensions[1].height = 35

    # Subtitle
    ws_test.merge_cells('A2:H2')
    sub_cell = ws_test['A2']
    sub_cell.value = 'Answer all 10 questions by selecting A, B, C, or D'
    sub_cell.font = Font(name='Arial', size=11, italic=True, color='666666')
    sub_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_test.row_dimensions[2].height = 25

    # Column headers for test area (row 4)
    test_headers = ['#', 'Question', 'Option A', 'Option B', 'Option C', 'Option D', 'Your Answer', 'Result']
    test_header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    test_header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')

    for col, h in enumerate(test_headers, 1):
        cell = ws_test.cell(row=4, column=col, value=h)
        cell.font = test_header_font
        cell.fill = test_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Question rows 5-14 (10 questions) — numbered but otherwise empty in initial
    for r in range(5, 15):
        q_num = r - 4
        ws_test.cell(row=r, column=1, value=q_num).alignment = center_align
        ws_test.cell(row=r, column=1).font = Font(name='Arial', size=10, bold=True)
        ws_test.cell(row=r, column=1).border = thin_border
        for c in range(2, 9):
            cell = ws_test.cell(row=r, column=c)
            cell.border = thin_border
            if c == 7:
                cell.alignment = center_align
                cell.font = Font(name='Arial', size=11, bold=True)
            elif c == 8:
                cell.alignment = center_align

    # Score section (row 16 onward) — labels only, no formulas
    ws_test.merge_cells('A16:F16')
    ws_test['A16'].value = 'Score Summary'
    ws_test['A16'].font = Font(name='Arial', size=13, bold=True, color='2F5496')
    ws_test['A16'].alignment = Alignment(horizontal='left', vertical='center')
    ws_test.row_dimensions[16].height = 28

    score_labels = [
        (17, 'Total Correct:'),
        (18, 'Total Questions:'),
        (19, 'Percentage:'),
        (20, 'Letter Grade:'),
    ]
    label_font = Font(name='Arial', size=11, bold=True)
    for row_num, label in score_labels:
        cell = ws_test.cell(row=row_num, column=6, value=label)
        cell.font = label_font
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = thin_border
        # Column G will hold formulas in golden, empty in initial
        ws_test.cell(row=row_num, column=7).border = thin_border
        ws_test.cell(row=row_num, column=7).alignment = center_align

    # Pre-fill "Total Questions" with static value 10
    ws_test.cell(row=18, column=7, value=10).font = Font(name='Arial', size=11)

    # Column widths for Test sheet
    ws_test.column_dimensions['A'].width = 5
    ws_test.column_dimensions['B'].width = 50
    ws_test.column_dimensions['C'].width = 20
    ws_test.column_dimensions['D'].width = 20
    ws_test.column_dimensions['E'].width = 20
    ws_test.column_dimensions['F'].width = 20
    ws_test.column_dimensions['G'].width = 14
    ws_test.column_dimensions['H'].width = 10

    # Set Test as the active sheet
    wb.active = wb.sheetnames.index('Test')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
