"""
Reward Script: Quiz/Test Generator and Grader
Task ID: calc_wf_058
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20): Bank sheet is hidden
  Component 2 (0.25): Question formulas in B5:B14 use INDEX/RANDBETWEEN to pull from Bank
  Component 3 (0.15): Grading formulas in H5:H14 compare student answer to correct answer
  Component 4 (0.10): Score summary formulas (total correct, percentage, letter grade)
  Component 5 (0.10): Data validation on G5:G14 for A/B/C/D
  Component 6 (0.10): Conditional formatting on result column (green/red)
  Component 7 (0.10): Option formulas in C5:F14 pull from Bank via INDEX/MATCH
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_058'


def _check_conditional_formatting(ws):
    """Check if conditional formatting exists on H column with rules for 1 and 0."""
    for cf in ws.conditional_formatting:
        range_str = str(cf).upper()
        if 'H' not in range_str:
            continue
        correct_rule_count = 0
        incorrect_rule_count = 0
        for rule in cf.rules:
            formula_str = str(rule.formula) if rule.formula else ''
            # Check for rule matching value 1 (correct) or 0 (incorrect)
            if '1' in formula_str:
                correct_rule_count += 1
            if '0' in formula_str:
                incorrect_rule_count += 1
        if correct_rule_count > 0 and incorrect_rule_count > 0:
            return 1  # both rules present
    return 0  # not found


def verify_task(file_path):
    """
    Verify quiz/test generator and grader task completion.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Both Bank and Test sheets must exist
    if 'Bank' not in wb.sheetnames or 'Test' not in wb.sheetnames:
        print(f"CRITICAL: Missing required sheets. Found: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws_bank = wb['Bank']
    ws_test = wb['Test']

    # Component 1: Bank sheet is hidden (0.20 points)
    # Initial: Bank is visible. Golden: Bank is hidden.
    try:
        if ws_bank.sheet_state == 'hidden' or ws_bank.sheet_state == 'veryHidden':
            print(f"PASS: Component 1 — Bank sheet is hidden (state={ws_bank.sheet_state}) (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 — Bank sheet state is '{ws_bank.sheet_state}', expected 'hidden'")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Question formulas in B5:B14 use INDEX with RANDBETWEEN to pull from Bank (0.25 points)
    # Initial: B5:B14 are all None. Golden: each has =INDEX(Bank!A:A,RANDBETWEEN(2,31))
    try:
        formula_count = 0
        for row in range(5, 15):
            val = ws_test.cell(row=row, column=2).value  # column B
            if val is not None and isinstance(val, str):
                val_upper = val.upper().replace(" ", "")
                if 'INDEX' in val_upper and 'BANK!' in val_upper and 'RANDBETWEEN' in val_upper:
                    formula_count += 1
        if formula_count == 10:
            print(f"PASS: Component 2 — All 10 question cells have INDEX/RANDBETWEEN formulas (0.25 pts)")
            total_score += 0.25
        elif formula_count >= 5:
            partial = round(0.25 * formula_count / 10, 2)
            print(f"PARTIAL: Component 2 — {formula_count}/10 question formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {formula_count}/10 question cells have INDEX/RANDBETWEEN formulas")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Grading formulas in H5:H14 compare student answer to correct answer (0.15 points)
    # Initial: H5:H14 are all None. Golden: each has =IF(G5="","",IF(G5=INDEX(Bank!F:F,MATCH(...)),1,0))
    try:
        grade_count = 0
        for row in range(5, 15):
            val = ws_test.cell(row=row, column=8).value  # column H
            if val is not None and isinstance(val, str):
                val_upper = val.upper().replace(" ", "")
                # Must reference both the student answer column (G) and Bank's correct answer
                if 'IF(' in val_upper and 'BANK!' in val_upper and ('MATCH' in val_upper or 'INDEX' in val_upper):
                    grade_count += 1
        if grade_count == 10:
            print(f"PASS: Component 3 — All 10 grading formulas present in H5:H14 (0.15 pts)")
            total_score += 0.15
        elif grade_count >= 5:
            partial = round(0.15 * grade_count / 10, 2)
            print(f"PARTIAL: Component 3 — {grade_count}/10 grading formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {grade_count}/10 grading formulas found in H5:H14")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Score summary formulas — total, percentage, letter grade (0.10 points)
    # Initial: G17, G19, G20 are all None. Golden: G17=SUM, G19=percentage, G20=IF-based grade
    try:
        summary_pts = 0.0
        # Check G17 for SUM formula (total correct)
        g17 = ws_test.cell(row=17, column=7).value  # G17
        if g17 is not None and isinstance(g17, str) and 'SUM' in g17.upper():
            summary_pts += 0.033
            print(f"  G17 SUM formula found: {g17}")

        # Check G19 for percentage formula
        g19 = ws_test.cell(row=19, column=7).value  # G19
        if g19 is not None and isinstance(g19, str) and '=' in g19:
            summary_pts += 0.033
            print(f"  G19 percentage formula found: {g19}")

        # Check G20 for letter grade formula (nested IF)
        g20 = ws_test.cell(row=20, column=7).value  # G20
        if g20 is not None and isinstance(g20, str) and 'IF' in g20.upper():
            summary_pts += 0.034
            print(f"  G20 letter grade formula found: {g20}")

        if summary_pts > 0.09:
            print(f"PASS: Component 4 — All 3 score summary formulas present (0.10 pts)")
            total_score += 0.10
        elif summary_pts > 0:
            total_score += round(summary_pts, 2)
            print(f"PARTIAL: Component 4 — Some score summary formulas present ({round(summary_pts, 2)} pts)")
        else:
            print(f"FAIL: Component 4 — No score summary formulas found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Data validation on G5:G14 for A/B/C/D input (0.10 points)
    # Initial: No data validations. Golden: list validation "A,B,C,D" on G5:G14
    try:
        dv_match_count = sum(
            1 for dv in ws_test.data_validations.dataValidation
            if dv.type == 'list'
            and all(ch in str(dv.formula1).upper() for ch in ['A', 'B', 'C', 'D'])
            and 'G' in str(dv.sqref).upper()
        )
        if dv_match_count > 0:
            print(f"PASS: Component 5 — Data validation for A/B/C/D found on answer cells (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 — No A/B/C/D list data validation found on G column")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Conditional formatting on result column for green/red (0.10 points)
    # Initial: No conditional formatting. Golden: H5:H14 has rules for =1 (green) and =0 (red)
    try:
        cf_pass = _check_conditional_formatting(ws_test)
        if cf_pass:
            print(f"PASS: Component 6 — Conditional formatting with green/red rules on result column (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 6 — No conditional formatting with correct/incorrect rules found on H column")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: Option formulas in C5:F14 pull options from Bank via INDEX/MATCH (0.10 points)
    # Initial: C5:F14 are all None. Golden: each has INDEX(Bank!...,MATCH(...))
    try:
        option_formula_count = 0
        for row in range(5, 15):
            for col in range(3, 7):  # columns C, D, E, F
                val = ws_test.cell(row=row, column=col).value
                if val is not None and isinstance(val, str):
                    val_upper = val.upper().replace(" ", "")
                    if 'INDEX' in val_upper and 'BANK!' in val_upper and 'MATCH' in val_upper:
                        option_formula_count += 1
        # 10 rows x 4 columns = 40 cells expected
        if option_formula_count >= 38:  # allow small tolerance
            print(f"PASS: Component 7 — {option_formula_count}/40 option formulas found in C5:F14 (0.10 pts)")
            total_score += 0.10
        elif option_formula_count >= 20:
            partial = round(0.10 * option_formula_count / 40, 2)
            print(f"PARTIAL: Component 7 — {option_formula_count}/40 option formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 7 — Only {option_formula_count}/40 option formulas found in C5:F14")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
