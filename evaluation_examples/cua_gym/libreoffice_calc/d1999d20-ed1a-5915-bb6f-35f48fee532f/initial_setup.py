"""
Initial Setup: Chart data series accidentally includes header row
Task ID: calc_tbl_063
Domain: libreoffice_calc

Creates a spreadsheet with monthly revenue data and a bar chart where
the data range includes the header row (A1:B13) WITHOUT titles_from_data,
causing 'Month' to appear as an X-axis label and 'Revenue' as a zero data point.
"""

import os
import shlex
import subprocess
import time

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_063'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    # --- Header row ---
    ws["A1"] = "Month"
    ws["B1"] = "Revenue"

    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    for cell in [ws["A1"], ws["B1"]]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    # --- Monthly data (rows 2-13) ---
    data = [
        ("January",   18450),
        ("February",  21300),
        ("March",     19875),
        ("April",     24600),
        ("May",       27150),
        ("June",      25800),
        ("July",      29400),
        ("August",    31200),
        ("September", 28650),
        ("October",   33100),
        ("November",  35750),
        ("December",  38200),
    ]

    for r, (month, revenue) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=month)
        ws.cell(row=r, column=2, value=revenue)
        ws.cell(row=r, column=2).number_format = '#,##0'

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14

    # --- Bar chart with BUGGY data range (includes header as data) ---
    chart = BarChart()
    chart.type = "col"
    chart.title = "Monthly Revenue"
    chart.y_axis.title = "Revenue ($)"
    chart.x_axis.title = "Month"
    chart.style = 10

    # BUG: data reference starts from row 1 WITHOUT titles_from_data=True
    # This causes 'Revenue' header to be plotted as a zero-value data point
    # and 'Month' to appear as the first X-axis category label
    data_ref = Reference(ws, min_col=2, min_row=1, max_row=13)
    cats_ref = Reference(ws, min_col=1, min_row=1, max_row=13)
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)

    chart.width = 18
    chart.height = 12

    ws.add_chart(chart, "D2")

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
