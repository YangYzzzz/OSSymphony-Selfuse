"""
Reward Script: Cash Flow Projection Spreadsheet
Task ID: calc_grs_039
Domain: libreoffice_calc
Scoring:
  - Total Revenue formulas (0.15)
  - Total Expenses formulas (0.15)
  - Net Cash Flow formulas (0.15)
  - Ending Cash Balance formulas (0.15)
  - Beginning Cash Balance linking (0.10)
  - Buffer Months formulas (0.10)
  - Chart present (0.10)
  - Conditional formatting on Ending Cash Balance (0.10)
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_039'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Find the main sheet (could be named differently by agent)
    ws = None
    for sn in wb.sheetnames:
        if 'cash' in sn.lower() or 'flow' in sn.lower() or 'projection' in sn.lower():
            ws = wb[sn]
            break
    if ws is None:
        ws = wb.worksheets[0]

    # Helper: check if a cell contains a formula (string starting with =)
    def is_formula(cell_val):
        return isinstance(cell_val, str) and cell_val.startswith('=')

    def formula_contains(cell_val, *keywords):
        """Check if formula contains all given keywords (case-insensitive)."""
        if not is_formula(cell_val):
            return False
        upper = cell_val.upper().replace(' ', '')
        return all(kw.upper().replace(' ', '') in upper for kw in keywords)

    # =========================================================================
    # Component 1: Total Revenue formulas in row 8, B8:M8 (0.15 pts)
    # These should be SUM formulas summing revenue rows (rows 4-7)
    # Initial has NO formulas here, golden has =SUM(B4:B7) etc.
    # =========================================================================
    try:
        rev_formula_count = 0
        for col in range(2, 14):  # B=2 to M=13
            val = ws.cell(row=8, column=col).value
            if is_formula(val) and formula_contains(val, 'SUM'):
                rev_formula_count += 1
        if rev_formula_count >= 10:
            print(f"PASS: Component 1 -- Total Revenue SUM formulas found in {rev_formula_count}/12 cells (0.15 pts)")
            total_score += 0.15
        elif rev_formula_count >= 6:
            partial = 0.15 * (rev_formula_count / 12)
            print(f"PARTIAL: Component 1 -- Total Revenue SUM formulas: {rev_formula_count}/12 ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 -- Total Revenue SUM formulas: only {rev_formula_count}/12 found")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # =========================================================================
    # Component 2: Total Expenses formulas in row 18, B18:M18 (0.15 pts)
    # Should be SUM formulas summing expense rows (rows 10-17)
    # Initial has NO formulas here.
    # =========================================================================
    try:
        exp_formula_count = 0
        for col in range(2, 14):
            val = ws.cell(row=18, column=col).value
            if is_formula(val) and formula_contains(val, 'SUM'):
                exp_formula_count += 1
        if exp_formula_count >= 10:
            print(f"PASS: Component 2 -- Total Expenses SUM formulas found in {exp_formula_count}/12 cells (0.15 pts)")
            total_score += 0.15
        elif exp_formula_count >= 6:
            partial = 0.15 * (exp_formula_count / 12)
            print(f"PARTIAL: Component 2 -- Total Expenses SUM formulas: {exp_formula_count}/12 ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- Total Expenses SUM formulas: only {exp_formula_count}/12 found")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # =========================================================================
    # Component 3: Net Cash Flow formulas in row 20, B20:M20 (0.15 pts)
    # Should be formulas computing Revenue - Expenses (e.g. =B8-B18)
    # Initial has NO formulas here.
    # =========================================================================
    try:
        ncf_formula_count = 0
        for col in range(2, 14):
            val = ws.cell(row=20, column=col).value
            # Should be a subtraction formula referencing total revenue and total expenses
            if is_formula(val):
                # Accept any formula that subtracts (contains minus sign)
                if '-' in val:
                    ncf_formula_count += 1
        if ncf_formula_count >= 10:
            print(f"PASS: Component 3 -- Net Cash Flow formulas found in {ncf_formula_count}/12 cells (0.15 pts)")
            total_score += 0.15
        elif ncf_formula_count >= 6:
            partial = 0.15 * (ncf_formula_count / 12)
            print(f"PARTIAL: Component 3 -- Net Cash Flow formulas: {ncf_formula_count}/12 ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 -- Net Cash Flow formulas: only {ncf_formula_count}/12 found")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # =========================================================================
    # Component 4: Ending Cash Balance formulas in row 21, B21:M21 (0.15 pts)
    # Should compute Beginning + Net Cash Flow (e.g. =B2+B20)
    # Initial has NO formulas here.
    # =========================================================================
    try:
        ecb_formula_count = 0
        for col in range(2, 14):
            val = ws.cell(row=21, column=col).value
            if is_formula(val):
                # Should contain addition (beginning balance + net cash flow)
                if '+' in val:
                    ecb_formula_count += 1
        if ecb_formula_count >= 10:
            print(f"PASS: Component 4 -- Ending Cash Balance formulas found in {ecb_formula_count}/12 cells (0.15 pts)")
            total_score += 0.15
        elif ecb_formula_count >= 6:
            partial = 0.15 * (ecb_formula_count / 12)
            print(f"PARTIAL: Component 4 -- Ending Cash Balance formulas: {ecb_formula_count}/12 ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 -- Ending Cash Balance formulas: only {ecb_formula_count}/12 found")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # =========================================================================
    # Component 5: Beginning Cash Balance linking C2:M2 (0.10 pts)
    # Each month's beginning balance = previous month's ending balance.
    # e.g. C2=B21, D2=C21, etc.
    # Initial has only B2=50000 with C2:M2 empty.
    # =========================================================================
    try:
        bcb_link_count = 0
        for col in range(3, 14):  # C=3 to M=13
            val = ws.cell(row=2, column=col).value
            if is_formula(val):
                # Should reference the previous month's ending cash balance (row 21)
                bcb_link_count += 1
        if bcb_link_count >= 9:
            print(f"PASS: Component 5 -- Beginning Cash Balance linking formulas: {bcb_link_count}/11 (0.10 pts)")
            total_score += 0.10
        elif bcb_link_count >= 5:
            partial = 0.10 * (bcb_link_count / 11)
            print(f"PARTIAL: Component 5 -- Beginning Cash Balance linking: {bcb_link_count}/11 ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 -- Beginning Cash Balance linking: only {bcb_link_count}/11 found")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    # =========================================================================
    # Component 6: Buffer Months formulas in row 23, B23:M23 (0.10 pts)
    # Should compute ending balance / expenses (e.g. =B21/B18)
    # Initial has NO formulas here.
    # =========================================================================
    try:
        buf_formula_count = 0
        for col in range(2, 14):
            val = ws.cell(row=23, column=col).value
            if is_formula(val):
                # Should be a division formula
                if '/' in val:
                    buf_formula_count += 1
        if buf_formula_count >= 10:
            print(f"PASS: Component 6 -- Buffer Months formulas found in {buf_formula_count}/12 cells (0.10 pts)")
            total_score += 0.10
        elif buf_formula_count >= 6:
            partial = 0.10 * (buf_formula_count / 12)
            print(f"PARTIAL: Component 6 -- Buffer Months formulas: {buf_formula_count}/12 ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 6 -- Buffer Months formulas: only {buf_formula_count}/12 found")
    except Exception as e:
        print(f"ERROR: Component 6 -- {e}")

    # =========================================================================
    # Component 7: Chart present (0.10 pts)
    # Golden has a LineChart titled "Cash Balance Trajectory" with 1 series.
    # Initial has NO charts.
    # =========================================================================
    try:
        charts = ws._charts
        if len(charts) >= 1:
            print(f"PASS: Component 7 -- Chart found ({len(charts)} chart(s)) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 7 -- No charts found in worksheet")
    except Exception as e:
        print(f"ERROR: Component 7 -- {e}")

    # =========================================================================
    # Component 8: Conditional formatting on Ending Cash Balance row (0.10 pts)
    # Golden has CF on B21:M21: cellIs lessThan 10000 with red fill (FFFF0000).
    # Initial has NO conditional formatting.
    # =========================================================================
    try:
        cf_matches = [
            cf for cf in ws.conditional_formatting
            if '21' in str(cf)
            and any(r.type in ('cellIs', 'expression') for r in cf.rules)
        ]
        if len(cf_matches) >= 1:
            print(f"PASS: Component 8 -- Conditional formatting found on Ending Cash Balance row (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 8 -- No conditional formatting found on Ending Cash Balance row 21")
    except Exception as e:
        print(f"ERROR: Component 8 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved LibreOffice edits before verification
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
persist_app_state()

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
