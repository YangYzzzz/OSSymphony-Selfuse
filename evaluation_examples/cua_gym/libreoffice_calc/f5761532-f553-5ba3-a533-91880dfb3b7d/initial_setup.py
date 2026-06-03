"""
Initial Setup: Cash flow projection spreadsheet for a small SaaS business
Task ID: calc_grs_039
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_039'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Cash Flow Projection ---
    ws = wb.active
    ws.title = "Cash Flow Projection"

    # Months
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    # ---- Row 1: Headers ----
    ws.cell(row=1, column=1, value="Category")
    for i, m in enumerate(months):
        ws.cell(row=1, column=i + 2, value=m)

    # Style headers
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    for col in range(1, 14):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # ---- Row 2: Beginning Cash Balance ----
    ws.cell(row=2, column=1, value="Beginning Cash Balance")
    ws.cell(row=2, column=2, value=50000)  # Jan starting balance
    # Feb-Dec: leave empty (task is to link these via cell references)

    # ---- Revenue Section ----
    # Row 3: Section header
    ws.cell(row=3, column=1, value="REVENUE")
    ws.cell(row=3, column=1).font = Font(bold=True, color="1F7A1F")

    # Revenue data - realistic SaaS startup numbers
    revenue_data = {
        "Products": [12500, 13200, 11800, 14500, 15200, 16800, 17500, 18200, 19000, 20500, 21200, 22800],
        "Services": [8500, 7200, 9500, 8800, 10200, 11500, 9800, 12000, 11200, 10800, 13500, 14200],
        "Subscriptions": [22000, 23100, 24200, 25400, 26700, 28000, 29400, 30900, 32400, 34000, 35700, 37500],
        "Other": [1500, 800, 2200, 1000, 3500, 1200, 900, 2800, 1500, 2000, 1100, 3000],
    }

    rev_row = 4
    for label, values in revenue_data.items():
        ws.cell(row=rev_row, column=1, value=label)
        for i, val in enumerate(values):
            ws.cell(row=rev_row, column=i + 2, value=val)
        rev_row += 1

    # Row 8: Total Revenue (leave empty - task requires formulas)
    ws.cell(row=8, column=1, value="Total Revenue")
    ws.cell(row=8, column=1).font = Font(bold=True)

    # ---- Expense Section ----
    # Row 9: Section header
    ws.cell(row=9, column=1, value="EXPENSES")
    ws.cell(row=9, column=1).font = Font(bold=True, color="CC0000")

    expense_data = {
        "Salaries": [35000, 35000, 35000, 36500, 36500, 36500, 38000, 38000, 38000, 40000, 40000, 40000],
        "Rent": [4500, 4500, 4500, 4500, 4500, 4500, 4500, 4500, 4500, 4500, 4500, 4500],
        "Marketing": [5200, 6800, 4500, 7200, 8500, 5800, 9200, 6500, 7800, 8200, 10500, 12000],
        "Software": [2800, 2800, 2800, 2800, 2800, 3200, 3200, 3200, 3200, 3200, 3200, 3200],
        "COGS": [3800, 4100, 3600, 4400, 4600, 5100, 5300, 5500, 5800, 6200, 6400, 6900],
        "Utilities": [850, 850, 850, 900, 900, 1100, 1100, 1100, 900, 900, 850, 850],
        "Insurance": [1200, 1200, 1200, 1200, 1200, 1200, 1200, 1200, 1200, 1200, 1200, 1200],
        "Misc": [1500, 800, 2200, 1000, 3500, 1200, 900, 2800, 1500, 600, 1100, 2000],
    }

    exp_row = 10
    for label, values in expense_data.items():
        ws.cell(row=exp_row, column=1, value=label)
        for i, val in enumerate(values):
            ws.cell(row=exp_row, column=i + 2, value=val)
        exp_row += 1

    # Row 18: Total Expenses (leave empty - task requires formulas)
    ws.cell(row=18, column=1, value="Total Expenses")
    ws.cell(row=18, column=1).font = Font(bold=True)

    # Row 19: blank separator
    # Row 20: Net Cash Flow (leave empty - task requires formula)
    ws.cell(row=20, column=1, value="Net Cash Flow")
    ws.cell(row=20, column=1).font = Font(bold=True)

    # Row 21: Ending Cash Balance (leave empty - task requires formula)
    ws.cell(row=21, column=1, value="Ending Cash Balance")
    ws.cell(row=21, column=1).font = Font(bold=True)

    # Row 23: Buffer Months label (leave empty - task requires calculation)
    ws.cell(row=23, column=1, value="Buffer Months")
    ws.cell(row=23, column=1).font = Font(bold=True)

    # ---- Format currency columns ----
    currency_fmt = '$#,##0'
    for row in range(2, 24):
        for col in range(2, 14):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None and isinstance(cell.value, (int, float)):
                cell.number_format = currency_fmt

    # ---- Column widths ----
    ws.column_dimensions['A'].width = 25
    for col_letter in ['B','C','D','E','F','G','H','I','J','K','L','M']:
        ws.column_dimensions[col_letter].width = 14

    # Freeze panes: freeze row 1 and column A
    ws.freeze_panes = "B2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
