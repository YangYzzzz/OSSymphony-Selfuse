"""
Reward Script: Export specific sheets to separate CSV files
Task ID: calc_gsi_083
Domain: libreoffice_calc
Scoring: 6 components (one per CSV file), each worth ~0.167 points.
         Each component checks: file exists, correct header, correct row count,
         and data matches the corresponding xlsx sheet.
"""

import os
import csv
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_083'

# Mapping from expected CSV filename to the xlsx sheet name
CSV_SHEET_MAP = {
    'jan_data.csv': 'January',
    'feb_data.csv': 'February',
    'mar_data.csv': 'March',
    'apr_data.csv': 'April',
    'may_data.csv': 'May',
    'jun_data.csv': 'June',
}

EXPECTED_HEADER = ['Date', 'Product', 'Region', 'Units Sold', 'Revenue', 'Cost']
POINTS_PER_FILE = 1.0 / 6.0  # ~0.1667


def normalize_cell(value):
    """Convert an openpyxl cell value to the string form expected in CSV."""
    if value is None:
        return ''
    return str(value)


def check_csv_data_matches(ws, csv_rows):
    """Compare CSV data rows against xlsx sheet rows.
    Returns None if all match, or a string describing the first mismatch."""
    for r_idx in range(1, len(csv_rows)):  # skip header
        xlsx_row_vals = [
            normalize_cell(cell.value)
            for cell in list(ws.iter_rows(
                min_row=r_idx + 1, max_row=r_idx + 1,
                min_col=1, max_col=ws.max_column
            ))[0]
        ]
        csv_row_vals = csv_rows[r_idx]

        if len(xlsx_row_vals) != len(csv_row_vals):
            return f"row {r_idx+1}: col count differs"

        for c_idx in range(len(xlsx_row_vals)):
            x_val = xlsx_row_vals[c_idx]
            c_val = csv_row_vals[c_idx]
            if x_val != c_val:
                # Try numeric comparison for floating point tolerance
                try:
                    if abs(float(x_val) - float(c_val)) < 0.01:
                        continue
                except (ValueError, TypeError):
                    pass
                return f"row {r_idx+1}, col {c_idx+1}: xlsx='{x_val}' csv='{c_val}'"
    return None  # all match


def verify_task(xlsx_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: xlsx must exist and be loadable
    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load workbook {xlsx_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    for csv_name, sheet_name in CSV_SHEET_MAP.items():
        csv_path = os.path.join(WORKDIR, csv_name)

        # Component: <csv_name> exists, has correct header, row count, and data
        try:
            # Check 1: File exists
            if not os.path.exists(csv_path):
                print(f"FAIL: {csv_name} — file does not exist")
                continue

            # Check 2: Read CSV
            with open(csv_path, newline='') as f:
                reader = csv.reader(f)
                csv_rows = list(reader)

            if len(csv_rows) < 2:
                print(f"FAIL: {csv_name} — file is empty or has no data rows (rows={len(csv_rows)})")
                continue

            # Check 3: Header matches expected columns
            csv_header = csv_rows[0]
            header_ok = (csv_header == EXPECTED_HEADER)
            if not header_ok:
                print(f"FAIL: {csv_name} — header mismatch: got {csv_header}")
                continue

            # Check 4: Corresponding xlsx sheet exists
            if sheet_name not in wb.sheetnames:
                print(f"FAIL: {csv_name} — sheet '{sheet_name}' missing from workbook")
                continue

            ws = wb[sheet_name]

            # Check 5: Row count matches (xlsx has header + data, csv should match)
            xlsx_row_count = ws.max_row  # includes header row
            csv_row_count = len(csv_rows)  # includes header row
            if csv_row_count != xlsx_row_count:
                print(f"FAIL: {csv_name} — row count mismatch: csv={csv_row_count}, xlsx={xlsx_row_count}")
                continue

            # Check 6: Verify data content matches (all rows)
            mismatch_detail = check_csv_data_matches(ws, csv_rows)
            if mismatch_detail is not None:
                print(f"FAIL: {csv_name} — data mismatch: {mismatch_detail}")
                continue

            # All checks passed for this CSV file
            if header_ok and csv_row_count == xlsx_row_count:
                print(f"PASS: {csv_name} — matches sheet '{sheet_name}' ({POINTS_PER_FILE:.4f} pts)")
                total_score += POINTS_PER_FILE

        except Exception as e:
            print(f"ERROR: {csv_name} — {e}")

    final_score = round(min(total_score, 1.0), 2)
    # Handle floating point: 6 * (1/6) should be 1.0
    if abs(final_score - 1.0) < 0.01:
        final_score = 1.0
    print(f"\nScore: {total_score:.4f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
