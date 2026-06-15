"""
Initial Setup: Export specific sheets to separate CSV files
Task ID: calc_gsi_083
Domain: libreoffice_calc

Creates a workbook with 6 monthly data sheets (January-June) containing
realistic sales data, then opens it in LibreOffice Calc.
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_083'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # Month configs: (sheet_name, month_num, year)
    months = [
        ("January", 1, 2025),
        ("February", 2, 2025),
        ("March", 3, 2025),
        ("April", 4, 2025),
        ("May", 5, 2025),
        ("June", 6, 2025),
    ]

    headers = ["Date", "Product", "Region", "Units Sold", "Revenue", "Cost"]

    products = [
        "Wireless Headphones", "USB-C Hub", "Mechanical Keyboard",
        "Monitor Stand", "Webcam Pro", "Desk Lamp", "Ergonomic Mouse",
        "Laptop Sleeve", "Cable Organizer", "Screen Protector",
        "Power Strip", "Phone Mount"
    ]

    regions = ["North", "South", "East", "West", "Central"]

    # Deterministic data generation
    import random
    rng = random.Random(42)

    for idx, (month_name, month_num, year) in enumerate(months):
        if idx == 0:
            ws = wb.active
            ws.title = month_name
        else:
            ws = wb.create_sheet(month_name)

        # Write headers
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

        # Generate 12 rows of data per month
        num_days_map = {1: 31, 2: 28, 3: 31, 4: 30, 5: 31, 6: 30}
        max_day = num_days_map[month_num]

        for r in range(2, 14):  # rows 2-13 = 12 data rows
            day = rng.randint(1, max_day)
            date_str = f"{year}-{month_num:02d}-{day:02d}"
            product = rng.choice(products)
            region = rng.choice(regions)
            units = rng.randint(5, 200)
            unit_price = round(rng.uniform(12.50, 189.99), 2)
            revenue = round(units * unit_price, 2)
            cost = round(revenue * rng.uniform(0.35, 0.65), 2)

            ws.cell(row=r, column=1, value=date_str)
            ws.cell(row=r, column=2, value=product)
            ws.cell(row=r, column=3, value=region)
            ws.cell(row=r, column=4, value=units)
            ws.cell(row=r, column=5, value=revenue)
            ws.cell(row=r, column=6, value=cost)

        # Set column widths for readability
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Open in LibreOffice Calc for GUI-ready state
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
