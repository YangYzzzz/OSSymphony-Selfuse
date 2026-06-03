"""
Reward Script: Stock Portfolio Tracker in LibreOffice Calc
Task ID: calc_grs_032
Domain: libreoffice_calc
Scoring:
  C1: Formulas in G/H/I columns (Purchase Value, Current Value, Gain/Loss) - 0.20
  C2: Formulas in J/K columns (Gain/Loss %, Portfolio Weight %) - 0.15
  C3: Data validation (sector dropdown on column C) - 0.10
  C4: Conditional formatting on Gain/Loss % column (4 rules) - 0.20
  C5: Summary section (Total Portfolio Value, Invested, Gain/Loss, Best/Worst) - 0.20
  C6: Pie chart present - 0.15
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_032'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Find the main sheet (should be first sheet or named 'Portfolio')
    ws = None
    for name in wb.sheetnames:
        if 'portfolio' in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb.worksheets[0]
    print(f"Using sheet: '{ws.title}'")

    # ----------------------------------------------------------------
    # Component 1: Formulas in G/H/I columns for rows 2-16 (0.20 pts)
    # Purchase Value = Shares * Purchase Price
    # Current Value = Shares * Current Price
    # Gain/Loss = Current Value - Purchase Value
    # These are EMPTY in initial_env, so scoring them is valid.
    # ----------------------------------------------------------------
    try:
        formula_count_ghi = 0
        expected_ghi = 15 * 3  # 15 rows, 3 columns
        for row in range(2, 17):
            g_val = ws.cell(row=row, column=7).value  # G = Purchase Value
            h_val = ws.cell(row=row, column=8).value  # H = Current Value
            i_val = ws.cell(row=row, column=9).value  # I = Gain/Loss
            for val in [g_val, h_val, i_val]:
                if isinstance(val, str) and val.startswith('='):
                    formula_count_ghi += 1
        ratio_ghi = formula_count_ghi / expected_ghi
        if ratio_ghi >= 0.8:
            print(f"PASS: Component 1 -- G/H/I formulas present: {formula_count_ghi}/{expected_ghi} (0.20 pts)")
            total_score += 0.20
        elif ratio_ghi >= 0.5:
            pts = round(0.10, 2)
            print(f"PARTIAL: Component 1 -- G/H/I formulas: {formula_count_ghi}/{expected_ghi} ({pts} pts)")
            total_score += pts
        else:
            print(f"FAIL: Component 1 -- G/H/I formulas: {formula_count_ghi}/{expected_ghi}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # ----------------------------------------------------------------
    # Component 2: Formulas in J/K columns for rows 2-16 (0.15 pts)
    # Gain/Loss % = Gain/Loss / Purchase Value
    # Portfolio Weight % = Current Value / Total Current Value
    # These are EMPTY in initial_env.
    # ----------------------------------------------------------------
    try:
        formula_count_jk = 0
        expected_jk = 15 * 2  # 15 rows, 2 columns
        for row in range(2, 17):
            j_val = ws.cell(row=row, column=10).value  # J = Gain/Loss %
            k_val = ws.cell(row=row, column=11).value  # K = Portfolio Weight %
            for val in [j_val, k_val]:
                if isinstance(val, str) and val.startswith('='):
                    formula_count_jk += 1
        ratio_jk = formula_count_jk / expected_jk
        if ratio_jk >= 0.8:
            print(f"PASS: Component 2 -- J/K formulas present: {formula_count_jk}/{expected_jk} (0.15 pts)")
            total_score += 0.15
        elif ratio_jk >= 0.5:
            pts = round(0.075, 3)
            print(f"PARTIAL: Component 2 -- J/K formulas: {formula_count_jk}/{expected_jk} ({pts} pts)")
            total_score += pts
        else:
            print(f"FAIL: Component 2 -- J/K formulas: {formula_count_jk}/{expected_jk}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # ----------------------------------------------------------------
    # Component 3: Data validation - sector dropdown on column C (0.10 pts)
    # Initial has NO data validations. Golden has a list validation.
    # ----------------------------------------------------------------
    try:
        dv_found = False
        if ws.data_validations and ws.data_validations.dataValidation:
            for dv in ws.data_validations.dataValidation:
                if dv.type == 'list' and dv.formula1:
                    # Check that the formula contains at least some sector names
                    sectors_in_formula = dv.formula1.lower()
                    known_sectors = ['technology', 'healthcare', 'financials', 'energy']
                    matches = sum(1 for s in known_sectors if s in sectors_in_formula)
                    if matches >= 3:
                        dv_found = True
                        break
        if dv_found:
            print(f"PASS: Component 3 -- Sector dropdown validation found (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 -- No sector dropdown validation found")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # ----------------------------------------------------------------
    # Component 4: Conditional formatting on Gain/Loss % column (0.20 pts)
    # Initial has NO conditional formatting. Golden has 4 rules on J2:J16.
    # We check: at least 3 CF rules that apply to column J range.
    # ----------------------------------------------------------------
    try:
        cf_rules_on_j = 0
        for cf in ws.conditional_formatting:
            cf_range_str = str(cf)
            # Check if the CF range involves column J (column 10)
            if 'J' in cf_range_str.upper():
                cf_rules_on_j += len(cf.rules)
        if cf_rules_on_j >= 4:
            print(f"PASS: Component 4 -- {cf_rules_on_j} CF rules on column J (0.20 pts)")
            total_score += 0.20
        elif cf_rules_on_j >= 2:
            pts = round(0.10, 2)
            print(f"PARTIAL: Component 4 -- {cf_rules_on_j} CF rules on column J ({pts} pts)")
            total_score += pts
        else:
            print(f"FAIL: Component 4 -- Only {cf_rules_on_j} CF rules on column J (need >= 4)")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # ----------------------------------------------------------------
    # Component 5: Summary section with key formulas (0.20 pts)
    # Initial has NO data below row 16. Golden has summary in rows 18-23.
    # We look for: Total Portfolio Value (SUM of H), Total Invested (SUM of G),
    # Total Gain/Loss (subtraction), Best/Worst Performer (MAX/MIN + INDEX/MATCH)
    # ----------------------------------------------------------------
    try:
        summary_items_found = 0

        # Search rows 17-40 for summary content
        summary_labels = {}
        summary_formulas = {}
        for row in range(17, 41):
            label = ws.cell(row=row, column=1).value
            value = ws.cell(row=row, column=2).value
            if label:
                summary_labels[row] = str(label).lower().strip()
            if value and isinstance(value, str) and value.startswith('='):
                summary_formulas[row] = value.upper().replace(' ', '')

        # Check for Total Portfolio Value (SUM of H or Current Value)
        for row, formula in summary_formulas.items():
            if 'SUM' in formula and 'H' in formula:
                summary_items_found += 1
                break

        # Check for Total Invested (SUM of G or Purchase Value)
        for row, formula in summary_formulas.items():
            if 'SUM' in formula and 'G' in formula:
                summary_items_found += 1
                break

        # Check for Total Gain/Loss (difference formula)
        for row, formula in summary_formulas.items():
            # Could be =B19-B20 or =SUM(I2:I16) or similar
            if ('-' in formula and 'SUM' not in formula) or ('SUM' in formula and 'I' in formula):
                summary_items_found += 1
                break

        # Check for Best/Worst Performer (MAX or MIN formula)
        max_found = False
        min_found = False
        for row, formula in summary_formulas.items():
            if 'MAX' in formula:
                max_found = True
            if 'MIN' in formula:
                min_found = True
        if max_found:
            summary_items_found += 1
        if min_found:
            summary_items_found += 1

        if summary_items_found >= 4:
            print(f"PASS: Component 5 -- Summary section: {summary_items_found}/5 items found (0.20 pts)")
            total_score += 0.20
        elif summary_items_found >= 2:
            pts = round(0.10, 2)
            print(f"PARTIAL: Component 5 -- Summary section: {summary_items_found}/5 items ({pts} pts)")
            total_score += pts
        else:
            print(f"FAIL: Component 5 -- Summary section: only {summary_items_found}/5 items found")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    # ----------------------------------------------------------------
    # Component 6: Pie chart present (0.15 pts)
    # Initial has 0 charts. Golden has 1 chart.
    # ----------------------------------------------------------------
    try:
        total_charts = 0
        for sn in wb.sheetnames:
            sheet = wb[sn]
            total_charts += len(sheet._charts)

        if total_charts >= 1:
            # Check if at least one is a pie chart
            pie_found = False
            for sn in wb.sheetnames:
                sheet = wb[sn]
                for chart in sheet._charts:
                    chart_type = type(chart).__name__.lower()
                    if 'pie' in chart_type:
                        pie_found = True
                        break
                if pie_found:
                    break

            if pie_found:
                print(f"PASS: Component 6 -- Pie chart found (0.15 pts)")
                total_score += 0.15
            else:
                # Chart exists but not a pie chart - partial credit
                pts = round(0.07, 2)
                print(f"PARTIAL: Component 6 -- Chart found but not pie type ({pts} pts)")
                total_score += pts
        else:
            print(f"FAIL: Component 6 -- No charts found")
    except Exception as e:
        print(f"ERROR: Component 6 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
