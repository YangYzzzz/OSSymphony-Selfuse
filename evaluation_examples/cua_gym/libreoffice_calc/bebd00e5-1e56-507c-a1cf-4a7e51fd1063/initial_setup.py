"""
Initial Setup: Stock Portfolio Tracker
Task ID: calc_grs_032
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_032'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Portfolio"

    # --- Headers ---
    headers = [
        "Ticker Symbol", "Company Name", "Sector", "Shares Owned",
        "Purchase Price", "Current Price", "Purchase Value",
        "Current Value", "Gain/Loss", "Gain/Loss %", "Portfolio Weight %"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # --- 15 Stock Holdings (realistic data, no formulas) ---
    stocks = [
        ["AAPL", "Apple Inc.", "Technology", 50, 142.50, 178.25],
        ["MSFT", "Microsoft Corp.", "Technology", 30, 285.00, 415.60],
        ["JNJ", "Johnson & Johnson", "Healthcare", 40, 165.30, 155.80],
        ["JPM", "JPMorgan Chase & Co.", "Financials", 25, 138.75, 196.40],
        ["XOM", "Exxon Mobil Corp.", "Energy", 60, 88.20, 105.50],
        ["PG", "Procter & Gamble Co.", "Consumer Staples", 35, 148.60, 162.35],
        ["AMZN", "Amazon.com Inc.", "Consumer Discretionary", 20, 132.40, 185.70],
        ["NEE", "NextEra Energy Inc.", "Utilities", 45, 78.50, 72.15],
        ["PLD", "Prologis Inc.", "Real Estate", 30, 125.40, 118.90],
        ["AMT", "American Tower Corp.", "Real Estate", 15, 215.80, 198.45],
        ["UNH", "UnitedHealth Group", "Healthcare", 10, 485.20, 542.30],
        ["V", "Visa Inc.", "Financials", 25, 228.90, 275.60],
        ["HD", "Home Depot Inc.", "Consumer Discretionary", 18, 305.40, 358.20],
        ["LIN", "Linde plc", "Technology", 22, 345.70, 412.85],
        ["CVX", "Chevron Corp.", "Energy", 35, 155.60, 162.40],
    ]

    for r, stock in enumerate(stocks, 2):
        for c, val in enumerate(stock, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths for readability
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 15
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 13
    ws.column_dimensions["K"].width = 17

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
