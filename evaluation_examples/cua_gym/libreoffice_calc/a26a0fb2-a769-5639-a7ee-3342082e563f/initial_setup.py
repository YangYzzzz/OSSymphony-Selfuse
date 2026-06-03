"""
Initial Setup: Make vs Buy Cost Analysis Model
Task ID: calc_ops_cost_analysis_make_vs_buy_071
Domain: libreoffice_calc

Creates a spreadsheet with component cost data for a make-vs-buy analysis.
The initial file has raw data only - no formulas for totals, no break-even column,
no cheaper option analysis, and no conditional formatting.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_cost_analysis_make_vs_buy_071'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'MakeVsBuy'

    # --- Header row 1: Component data columns A-G + Volume columns H-J ---
    header_font = Font(bold=True, name='Calibri', size=11)
    header_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center')
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Component data headers
    col_headers = [
        'Component',
        'Make: Material Cost/unit',
        'Make: Labor Cost/unit',
        'Make: Overhead/unit',
        'Make: Setup Cost',
        'Buy: Supplier Quote/unit',
        'Buy: Tooling/Setup',
    ]
    for col_idx, header in enumerate(col_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # Volume scenario headers (H1=1000, I1=5000, J1=10000)
    volume_fill = PatternFill(start_color='FFFCE4D6', end_color='FFFCE4D6', fill_type='solid')
    volumes = [1000, 5000, 10000]
    for idx, vol in enumerate(volumes, 8):
        cell = ws.cell(row=1, column=idx, value=vol)
        cell.font = header_font
        cell.fill = volume_fill
        cell.alignment = center_align
        cell.border = border

    # --- Component data rows 2-6 ---
    components = [
        # Component, Material/unit, Labor/unit, Overhead/unit, Setup Cost, Supplier Quote/unit, Tooling/Setup
        ('Hydraulic Pump Housing',    18.50,  12.75,  6.20,  4200.00,  42.00,  800.00),
        ('Gear Assembly – Type B',    22.30,  15.40,  7.85,  5800.00,  52.50,  1200.00),
        ('Valve Body – 3/4 inch',     11.60,   9.20,  4.30,  2900.00,  29.80,  500.00),
        ('Bracket – Mounting Plate',   6.40,   5.10,  2.95,  1500.00,  17.25,  300.00),
        ('Impeller – Centrifugal',    31.80,  20.60,  9.40,  7500.00,  68.00,  1800.00),
    ]

    data_font = Font(name='Calibri', size=11)
    currency_fmt = '#,##0.00'
    integer_fmt = '#,##0'

    for row_offset, comp in enumerate(components):
        row = row_offset + 2  # rows 2-6
        (name, mat, labor, overhead, setup, quote, tooling) = comp

        # A: Component name
        cell_a = ws.cell(row=row, column=1, value=name)
        cell_a.font = data_font
        cell_a.alignment = left_align
        cell_a.border = border

        # B: Material cost/unit
        cell_b = ws.cell(row=row, column=2, value=mat)
        cell_b.font = data_font
        cell_b.number_format = currency_fmt
        cell_b.alignment = center_align
        cell_b.border = border

        # C: Labor cost/unit
        cell_c = ws.cell(row=row, column=3, value=labor)
        cell_c.font = data_font
        cell_c.number_format = currency_fmt
        cell_c.alignment = center_align
        cell_c.border = border

        # D: Overhead/unit
        cell_d = ws.cell(row=row, column=4, value=overhead)
        cell_d.font = data_font
        cell_d.number_format = currency_fmt
        cell_d.alignment = center_align
        cell_d.border = border

        # E: Setup cost (Make)
        cell_e = ws.cell(row=row, column=5, value=setup)
        cell_e.font = data_font
        cell_e.number_format = currency_fmt
        cell_e.alignment = center_align
        cell_e.border = border

        # F: Supplier quote/unit
        cell_f = ws.cell(row=row, column=6, value=quote)
        cell_f.font = data_font
        cell_f.number_format = currency_fmt
        cell_f.alignment = center_align
        cell_f.border = border

        # G: Tooling/setup (Buy)
        cell_g = ws.cell(row=row, column=7, value=tooling)
        cell_g.font = data_font
        cell_g.number_format = currency_fmt
        cell_g.alignment = center_align
        cell_g.border = border

    # --- Section label rows ---
    section_font = Font(bold=True, name='Calibri', size=11)
    section_fill = PatternFill(start_color='FFE2EFDA', end_color='FFE2EFDA', fill_type='solid')

    # Row 7: "Make Total Cost at Volume" section label
    ws.cell(row=7, column=1, value='Make Total Cost at Volume').font = section_font
    ws.cell(row=7, column=1).fill = section_fill

    # Row 13: "Buy Total Cost at Volume" section label
    ws.cell(row=13, column=1, value='Buy Total Cost at Volume').font = section_font
    ws.cell(row=13, column=1).fill = section_fill

    # Row 19: "Cheaper Option" section label
    ws.cell(row=19, column=1, value='Cheaper Option (MAKE or BUY)').font = section_font
    ws.cell(row=19, column=1).fill = section_fill

    # Row 25: "Break-Even Quantity" section label
    ws.cell(row=25, column=1, value='Break-Even Quantity (units)').font = section_font
    ws.cell(row=25, column=1).fill = section_fill

    # Component name labels in rows 8-12, 14-18, 20-24, 26-30
    for row_offset, comp in enumerate(components):
        comp_name = comp[0]
        for section_start in [8, 14, 20, 26]:
            row = section_start + row_offset
            cell = ws.cell(row=row, column=1, value=comp_name)
            cell.font = data_font
            cell.alignment = left_align
            cell.border = border

    # Columns H-J headers repeated in rows 7, 13, 19 (volume labels)
    for section_row in [7, 13, 19]:
        for idx, vol in enumerate(volumes, 8):
            cell = ws.cell(row=section_row, column=idx, value=vol)
            cell.font = section_font
            cell.fill = section_fill
            cell.alignment = center_align
            cell.border = border

    # --- Column widths ---
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 24
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 16
    ws.column_dimensions['J'].width = 16
    ws.column_dimensions['K'].width = 24

    # Row heights
    ws.row_dimensions[1].height = 45

    # Freeze panes at B2
    ws.freeze_panes = 'B2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: MakeVsBuy')
    print('Rows 2-6: 5 components with cost data')
    print('Rows 8-12: Make Total Cost section (empty H:J)')
    print('Rows 14-18: Buy Total Cost section (empty H:J)')
    print('Rows 20-24: Cheaper Option section (empty H:J)')
    print('Rows 26-30: Break-Even Quantity section (empty)')
    print('NO formulas, NO conditional formatting in initial file')

create_initial()
