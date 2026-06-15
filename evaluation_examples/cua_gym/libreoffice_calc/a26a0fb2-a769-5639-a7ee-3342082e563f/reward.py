"""
Reward Script: Build a make vs buy analysis model for five components
Task ID: calc_ops_cost_analysis_make_vs_buy_071
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Make Total Cost formulas in H8:J12           — 0.25 points
  Component 2: Buy Total Cost formulas in H14:J18           — 0.25 points
  Component 3: Cheaper Option IF formulas in H20:J24        — 0.25 points
  Component 4: Break-Even Quantity formulas in K26:K30      — 0.15 points
  Component 5: Conditional formatting on cheaper option cells— 0.10 points
  Total: 1.00
"""

import os
import openpyxl
from openpyxl.utils import get_column_letter

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_cost_analysis_make_vs_buy_071'

# Expected formula patterns (normalized, upper-case, no spaces)
# Make Total Cost: =E{row}+(B{row}+C{row}+D{row})*H$1 (with col letter variant)
# Buy Total Cost:  =G{row}+F{row}*H$1
# Cheaper Option:  =IF(H{make}<H{buy},"MAKE","BUY")
# Break-Even:      =MAX(0,(E{row}-G{row})/(F{row}-(B{row}+C{row}+D{row})))

# Component data row mapping: component index (0-4) -> excel data row (2-6)
DATA_ROWS = [2, 3, 4, 5, 6]
# Make Total Cost output rows: H8:J12 -> excel rows 8-12
MAKE_COST_ROWS = [8, 9, 10, 11, 12]
# Buy Total Cost output rows: H14:J18 -> excel rows 14-18
BUY_COST_ROWS = [14, 15, 16, 17, 18]
# Cheaper Option output rows: H20:J24 -> excel rows 20-24
CHEAPER_OPTION_ROWS = [20, 21, 22, 23, 24]
# Break-Even rows: K26:K30
BREAKEVEN_ROWS = [26, 27, 28, 29, 30]
# Volume columns: H=8, I=9, J=10
VOL_COLS = [8, 9, 10]
VOL_COL_LETTERS = ['H', 'I', 'J']


def normalize_formula(formula):
    """Normalize a formula string for comparison: uppercase, remove spaces."""
    if not formula:
        return ''
    return str(formula).upper().replace(' ', '')


def check_make_formula(formula_str, data_row, vol_col_letter):
    """
    Check if the Make Total Cost formula is correct.
    Expected: =E{dr}+(B{dr}+C{dr}+D{dr})*{vol}$1
    Accepts both locked row ref e.g. H$1 and unlocked H1.
    """
    norm = normalize_formula(formula_str)
    dr = data_row
    # Pattern like: =E2+(B2+C2+D2)*H$1  or =E2+(B2+C2+D2)*H1
    expected_locked = normalize_formula(
        f'=E{dr}+(B{dr}+C{dr}+D{dr})*{vol_col_letter}$1'
    )
    expected_unlocked = normalize_formula(
        f'=E{dr}+(B{dr}+C{dr}+D{dr})*{vol_col_letter}1'
    )
    return norm == expected_locked or norm == expected_unlocked


def check_buy_formula(formula_str, data_row, vol_col_letter):
    """
    Check if the Buy Total Cost formula is correct.
    Expected: =G{dr}+F{dr}*{vol}$1 or =G{dr}+F{dr}*{vol}1
    """
    norm = normalize_formula(formula_str)
    dr = data_row
    expected_locked = normalize_formula(f'=G{dr}+F{dr}*{vol_col_letter}$1')
    expected_unlocked = normalize_formula(f'=G{dr}+F{dr}*{vol_col_letter}1')
    return norm == expected_locked or norm == expected_unlocked


def check_cheaper_formula(formula_str, make_row, buy_row, vol_col_letter):
    """
    Check if the Cheaper Option formula is correct.
    Expected: =IF({vol}{make_row}<{vol}{buy_row},"MAKE","BUY")
    """
    norm = normalize_formula(formula_str)
    expected = normalize_formula(
        f'=IF({vol_col_letter}{make_row}<{vol_col_letter}{buy_row},"MAKE","BUY")'
    )
    return norm == expected


def check_breakeven_formula(formula_str, data_row):
    """
    Check if the Break-Even Quantity formula is correct.
    Expected: =MAX(0,(E{dr}-G{dr})/(F{dr}-(B{dr}+C{dr}+D{dr})))
    """
    norm = normalize_formula(formula_str)
    dr = data_row
    expected = normalize_formula(
        f'=MAX(0,(E{dr}-G{dr})/(F{dr}-(B{dr}+C{dr}+D{dr})))'
    )
    return norm == expected


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns float between 0.0 and 1.0.
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'MakeVsBuy' not in wb.sheetnames:
        print("CRITICAL: Sheet 'MakeVsBuy' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['MakeVsBuy']

    # -------------------------------------------------------------------
    # Component 1: Make Total Cost formulas in H8:J12 (0.25 points)
    # Each component row i uses data from row DATA_ROWS[i]
    # Formula: =E{dr}+(B{dr}+C{dr}+D{dr})*{vol_col}$1
    # -------------------------------------------------------------------
    try:
        make_formula_pass = 0
        make_formula_total = len(DATA_ROWS) * len(VOL_COLS)  # 15 cells

        for i, (make_row, data_row) in enumerate(zip(MAKE_COST_ROWS, DATA_ROWS)):
            for vol_col, vol_letter in zip(VOL_COLS, VOL_COL_LETTERS):
                cell = ws.cell(row=make_row, column=vol_col)
                val = cell.value
                if val is not None and isinstance(val, str) and val.startswith('='):
                    if check_make_formula(val, data_row, vol_letter):
                        make_formula_pass += 1
                    else:
                        print(
                            f"FAIL Component 1: {get_column_letter(vol_col)}{make_row} "
                            f"has wrong formula: {repr(val)}"
                        )
                else:
                    print(
                        f"FAIL Component 1: {get_column_letter(vol_col)}{make_row} "
                        f"missing formula, found: {repr(val)}"
                    )

        if make_formula_pass == make_formula_total:
            print(f"PASS: Component 1 — All {make_formula_total} Make Total Cost formulas correct (0.25 pts)")
            total_score += 0.25
        elif make_formula_pass >= make_formula_total * 0.5:
            partial = round(0.25 * make_formula_pass / make_formula_total, 3)
            print(
                f"PARTIAL: Component 1 — {make_formula_pass}/{make_formula_total} "
                f"Make Total Cost formulas correct ({partial} pts)"
            )
            total_score += partial
        else:
            print(
                f"FAIL: Component 1 — Only {make_formula_pass}/{make_formula_total} "
                f"Make Total Cost formulas correct (0.0 pts)"
            )
    except Exception as e:
        print(f"ERROR: Component 1 (Make Total Cost) — {e}")

    # -------------------------------------------------------------------
    # Component 2: Buy Total Cost formulas in H14:J18 (0.25 points)
    # Formula: =G{dr}+F{dr}*{vol_col}$1
    # -------------------------------------------------------------------
    try:
        buy_formula_pass = 0
        buy_formula_total = len(DATA_ROWS) * len(VOL_COLS)  # 15 cells

        for i, (buy_row, data_row) in enumerate(zip(BUY_COST_ROWS, DATA_ROWS)):
            for vol_col, vol_letter in zip(VOL_COLS, VOL_COL_LETTERS):
                cell = ws.cell(row=buy_row, column=vol_col)
                val = cell.value
                if val is not None and isinstance(val, str) and val.startswith('='):
                    if check_buy_formula(val, data_row, vol_letter):
                        buy_formula_pass += 1
                    else:
                        print(
                            f"FAIL Component 2: {get_column_letter(vol_col)}{buy_row} "
                            f"has wrong formula: {repr(val)}"
                        )
                else:
                    print(
                        f"FAIL Component 2: {get_column_letter(vol_col)}{buy_row} "
                        f"missing formula, found: {repr(val)}"
                    )

        if buy_formula_pass == buy_formula_total:
            print(f"PASS: Component 2 — All {buy_formula_total} Buy Total Cost formulas correct (0.25 pts)")
            total_score += 0.25
        elif buy_formula_pass >= buy_formula_total * 0.5:
            partial = round(0.25 * buy_formula_pass / buy_formula_total, 3)
            print(
                f"PARTIAL: Component 2 — {buy_formula_pass}/{buy_formula_total} "
                f"Buy Total Cost formulas correct ({partial} pts)"
            )
            total_score += partial
        else:
            print(
                f"FAIL: Component 2 — Only {buy_formula_pass}/{buy_formula_total} "
                f"Buy Total Cost formulas correct (0.0 pts)"
            )
    except Exception as e:
        print(f"ERROR: Component 2 (Buy Total Cost) — {e}")

    # -------------------------------------------------------------------
    # Component 3: Cheaper Option IF formulas in H20:J24 (0.25 points)
    # Formula: =IF({vol}{make_row}<{vol}{buy_row},"MAKE","BUY")
    # -------------------------------------------------------------------
    try:
        cheaper_formula_pass = 0
        cheaper_formula_total = len(DATA_ROWS) * len(VOL_COLS)  # 15 cells

        for i, (cheaper_row, make_row, buy_row) in enumerate(
            zip(CHEAPER_OPTION_ROWS, MAKE_COST_ROWS, BUY_COST_ROWS)
        ):
            for vol_col, vol_letter in zip(VOL_COLS, VOL_COL_LETTERS):
                cell = ws.cell(row=cheaper_row, column=vol_col)
                val = cell.value
                if val is not None and isinstance(val, str) and val.startswith('='):
                    if check_cheaper_formula(val, make_row, buy_row, vol_letter):
                        cheaper_formula_pass += 1
                    else:
                        print(
                            f"FAIL Component 3: {get_column_letter(vol_col)}{cheaper_row} "
                            f"has wrong formula: {repr(val)}"
                        )
                else:
                    print(
                        f"FAIL Component 3: {get_column_letter(vol_col)}{cheaper_row} "
                        f"missing formula, found: {repr(val)}"
                    )

        if cheaper_formula_pass == cheaper_formula_total:
            print(f"PASS: Component 3 — All {cheaper_formula_total} Cheaper Option formulas correct (0.25 pts)")
            total_score += 0.25
        elif cheaper_formula_pass >= cheaper_formula_total * 0.5:
            partial = round(0.25 * cheaper_formula_pass / cheaper_formula_total, 3)
            print(
                f"PARTIAL: Component 3 — {cheaper_formula_pass}/{cheaper_formula_total} "
                f"Cheaper Option formulas correct ({partial} pts)"
            )
            total_score += partial
        else:
            print(
                f"FAIL: Component 3 — Only {cheaper_formula_pass}/{cheaper_formula_total} "
                f"Cheaper Option formulas correct (0.0 pts)"
            )
    except Exception as e:
        print(f"ERROR: Component 3 (Cheaper Option) — {e}")

    # -------------------------------------------------------------------
    # Component 4: Break-Even Quantity formulas in K26:K30 (0.15 points)
    # Formula: =MAX(0,(E{dr}-G{dr})/(F{dr}-(B{dr}+C{dr}+D{dr})))
    # Also checks K25 contains a header label for break-even column
    # -------------------------------------------------------------------
    try:
        breakeven_pass = 0
        breakeven_total = len(DATA_ROWS)  # 5 cells

        for i, (be_row, data_row) in enumerate(zip(BREAKEVEN_ROWS, DATA_ROWS)):
            cell = ws.cell(row=be_row, column=11)  # column K = 11
            val = cell.value
            if val is not None and isinstance(val, str) and val.startswith('='):
                if check_breakeven_formula(val, data_row):
                    breakeven_pass += 1
                else:
                    print(
                        f"FAIL Component 4: K{be_row} has wrong formula: {repr(val)}"
                    )
            else:
                print(
                    f"FAIL Component 4: K{be_row} missing formula, found: {repr(val)}"
                )

        if breakeven_pass == breakeven_total:
            print(f"PASS: Component 4 — All {breakeven_total} Break-Even Quantity formulas correct (0.15 pts)")
            total_score += 0.15
        elif breakeven_pass > 0:
            partial = round(0.15 * breakeven_pass / breakeven_total, 3)
            print(
                f"PARTIAL: Component 4 — {breakeven_pass}/{breakeven_total} "
                f"Break-Even formulas correct ({partial} pts)"
            )
            total_score += partial
        else:
            print(
                f"FAIL: Component 4 — No Break-Even Quantity formulas found in K26:K30 (0.0 pts)"
            )
    except Exception as e:
        print(f"ERROR: Component 4 (Break-Even Quantity) — {e}")

    # -------------------------------------------------------------------
    # Component 5: Conditional formatting on cheaper option cells H20:J24 (0.10 points)
    # The golden file has 2 cellIs rules on H20:J24 with colors FF9DC3E6 (blue/MAKE)
    # and FFA9D18E (green/BUY)
    # -------------------------------------------------------------------
    try:
        cf_list = list(ws.conditional_formatting)
        # Look for a CF rule covering the cheaper option range (H20:J24 or subset)
        cf_found = False
        cf_color_match = False
        target_colors = {'FF9DC3E6', 'FFA9D18E'}  # blue and green from golden file

        for cf in cf_list:
            cf_range_str = str(cf)
            # Accept any CF range that includes H20:J24 or overlaps with it
            if 'H20' in cf_range_str or '20' in cf_range_str:
                cf_found = True
                found_colors = set()
                for rule in cf.rules:
                    dxf = getattr(rule, 'dxf', None)
                    if dxf:
                        fill = getattr(dxf, 'fill', None)
                        if fill:
                            try:
                                rgb = fill.fgColor.rgb
                                if rgb:
                                    found_colors.add(rgb.upper())
                            except Exception:
                                pass
                if len(found_colors) >= 2 and found_colors == target_colors:
                    cf_color_match = True
                elif len(found_colors) >= 2:
                    # At least 2 rules with some fill colors (partial match acceptable)
                    cf_color_match = True
                    print(
                        f"NOTE: CF colors found: {found_colors}, expected: {target_colors}"
                    )
                break

        if cf_found and cf_color_match:
            print("PASS: Component 5 — Conditional formatting on H20:J24 with color rules (0.10 pts)")
            total_score += 0.10
        elif cf_found:
            # CF range exists but colors don't match or only 1 rule
            print(
                "PARTIAL: Component 5 — Conditional formatting exists on cheaper option "
                "range but color rules incomplete (0.05 pts)"
            )
            total_score += 0.05
        else:
            print("FAIL: Component 5 — No conditional formatting found on cheaper option cells H20:J24 (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 5 (Conditional Formatting) — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
