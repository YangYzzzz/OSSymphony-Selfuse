"""
Initial Setup: Faculty CV Researcher Tracker with missing values
Task ID: osworld_multi_apps_web_scholar_004
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_web_scholar_004'
OUTPUT_ODS = f'{WORKDIR}/cv_researchers.ods'
DESKTOP = f'{WORKDIR}/Desktop'
DESKTOP_ODS = f'{DESKTOP}/cv_researchers.ods'
TMP_XLSX = f'{WORKDIR}/cv_researchers_tmp.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    import subprocess as _sp
    _sp.run(['pip3', 'install', 'openpyxl', '-q'], check=False)
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Researchers'

    # Headers
    headers = ['Name', 'Institution', 'H_Index', 'Specialization', 'Best_Known_For']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Data rows — with specific blanks as described in the task
    # None represents blank cells that the agent must fill in
    data = [
        # Row 2: H_Index blank
        ['Fei-Fei Li',       'Stanford',              None,  'Computer Vision',            'ImageNet'],
        # Row 3: Specialization and Best_Known_For blank
        ['Jitendra Malik',   'UC Berkeley',           80,    None,                          None],
        # Row 4: H_Index blank
        ['Kaiming He',       'Meta AI',               None,  'Deep Learning',              'ResNet'],
        # Row 5: Specialization blank
        ['Devi Parikh',      'Georgia Tech/Meta',     39,    None,                          'Visual Q&A'],
        # Additional rows for realism (complete data)
        ['Yann LeCun',       'Meta AI / NYU',         173,   'Deep Learning',              'Convolutional Neural Networks'],
        ['Geoffrey Hinton',  'University of Toronto', 189,   'Neural Networks',            'Backpropagation'],
        ['Yoshua Bengio',    'Universite de Montreal',205,   'Deep Learning, NLP',         'Generative Models'],
        ['Andrew Ng',        'Stanford / Coursera',   124,   'Machine Learning',           'Online ML Education'],
        ['Ian Goodfellow',   'Apple',                 85,    'Generative Models',          'Generative Adversarial Networks'],
        ['Pieter Abbeel',    'UC Berkeley',           67,    'Reinforcement Learning',     'Imitation Learning'],
        ['Andrej Karpathy',  'Tesla / OpenAI',        55,    'Computer Vision, LLMs',      'Deep RL, Neural Language Models'],
        ['Ruslan Salakhutdinov', 'Carnegie Mellon',   78,    'Probabilistic Models',       'Restricted Boltzmann Machines'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Save as xlsx temporarily
    wb.save(TMP_XLSX)
    print(f'Temporary xlsx created: {TMP_XLSX}')

    # Ensure Desktop directory exists
    os.makedirs(DESKTOP, exist_ok=True)

    # Convert xlsx to ods using LibreOffice headless
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    result = subprocess.run(
        ['libreoffice', '--headless', '--convert-to', 'ods', '--outdir', WORKDIR, TMP_XLSX],
        env=env,
        capture_output=True,
        text=True,
        timeout=30,
    )
    print(f'LibreOffice convert stdout: {result.stdout}')
    print(f'LibreOffice convert stderr: {result.stderr}')

    # The converted file will be named cv_researchers_tmp.ods in WORKDIR
    converted = f'{WORKDIR}/cv_researchers_tmp.ods'
    if os.path.exists(converted):
        import shutil
        shutil.move(converted, OUTPUT_ODS)
        print(f'Moved ODS to: {OUTPUT_ODS}')
        # Copy to Desktop
        shutil.copy(OUTPUT_ODS, DESKTOP_ODS)
        print(f'Copied ODS to Desktop: {DESKTOP_ODS}')
    else:
        # Fallback: try direct name
        alt = f'{WORKDIR}/cv_researchers.ods'
        if not os.path.exists(alt):
            print(f'ERROR: conversion failed, converted file not found at {converted}')
            # Try to copy tmp xlsx as fallback with ods name
            import shutil
            shutil.copy(TMP_XLSX, OUTPUT_ODS)
            shutil.copy(TMP_XLSX, DESKTOP_ODS)
            print(f'Fallback: copied xlsx as ods')
        else:
            import shutil
            shutil.copy(alt, DESKTOP_ODS)
            print(f'Copied ODS to Desktop: {DESKTOP_ODS}')

    # Remove temp xlsx
    if os.path.exists(TMP_XLSX):
        os.remove(TMP_XLSX)

    # Verify the file exists on Desktop
    if os.path.exists(DESKTOP_ODS):
        print(f'SUCCESS: {DESKTOP_ODS} exists')
    else:
        print(f'WARNING: {DESKTOP_ODS} not found')

    # GUI-ready startup: open the file in LibreOffice Calc
    time.sleep(1.0)
    launch_gui(f'libreoffice --calc "{DESKTOP_ODS}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with cv_researchers.ods on DISPLAY=:0')


create_initial()
