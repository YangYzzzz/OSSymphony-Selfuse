"""
Reward Script: Group sheets Monday-Friday and type Hours/8:00 AM/5:00 PM in A1-A3
Task ID: calc_ps_084
Domain: libreoffice_calc
Scoring:
  Component 1 (0.34): A1='Hours' on all 5 weekday sheets
  Component 2 (0.33): A2='8:00 AM' on all 5 weekday sheets
  Component 3 (0.33): A3='5:00 PM' on all 5 weekday sheets
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_084'
EXPECTED_SHEETS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: all 5 weekday sheets must exist
    for sheet_name in EXPECTED_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"CRITICAL: Sheet '{sheet_name}' not found. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0

    # Component 1: A1='Hours' on all 5 sheets (0.34 points)
    try:
        matching = 0
        for sheet_name in EXPECTED_SHEETS:
            ws = wb[sheet_name]
            val = ws['A1'].value
            if val is not None and str(val).strip() == 'Hours':
                matching += 1
            else:
                print(f"FAIL: Component 1 — {sheet_name} A1 expected 'Hours', found: {val!r}")
        if matching == 5:
            print(f"PASS: Component 1 — A1='Hours' on all 5 sheets (0.34 pts)")
            total_score += 0.34
        elif matching > 0:
            partial = round(0.34 * matching / 5, 4)
            print(f"PARTIAL: Component 1 — A1='Hours' on {matching}/5 sheets ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — A1='Hours' on 0/5 sheets")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: A2='8:00 AM' on all 5 sheets (0.33 points)
    try:
        matching = 0
        for sheet_name in EXPECTED_SHEETS:
            ws = wb[sheet_name]
            val = ws['A2'].value
            if val is not None and str(val).strip() == '8:00 AM':
                matching += 1
            else:
                print(f"FAIL: Component 2 — {sheet_name} A2 expected '8:00 AM', found: {val!r}")
        if matching == 5:
            print(f"PASS: Component 2 — A2='8:00 AM' on all 5 sheets (0.33 pts)")
            total_score += 0.33
        elif matching > 0:
            partial = round(0.33 * matching / 5, 4)
            print(f"PARTIAL: Component 2 — A2='8:00 AM' on {matching}/5 sheets ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — A2='8:00 AM' on 0/5 sheets")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: A3='5:00 PM' on all 5 sheets (0.33 points)
    try:
        matching = 0
        for sheet_name in EXPECTED_SHEETS:
            ws = wb[sheet_name]
            val = ws['A3'].value
            if val is not None and str(val).strip() == '5:00 PM':
                matching += 1
            else:
                print(f"FAIL: Component 3 — {sheet_name} A3 expected '5:00 PM', found: {val!r}")
        if matching == 5:
            print(f"PASS: Component 3 — A3='5:00 PM' on all 5 sheets (0.33 pts)")
            total_score += 0.33
        elif matching > 0:
            partial = round(0.33 * matching / 5, 4)
            print(f"PARTIAL: Component 3 — A3='5:00 PM' on {matching}/5 sheets ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — A3='5:00 PM' on 0/5 sheets")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist any unsaved GUI state before verification
persist_app_state("libreoffice_calc")

# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
