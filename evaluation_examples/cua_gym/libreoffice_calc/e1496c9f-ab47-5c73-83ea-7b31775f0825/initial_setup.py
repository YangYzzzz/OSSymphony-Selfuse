"""
Initial Setup: Fix #NAME? error from mistyped VLOOKUP function
Task ID: calc_tbl_066
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_066'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Product Lookup ---
    ws = wb.active
    ws.title = "Product Lookup"

    # Headers
    headers = ["Product Code", "Category", "Unit Price", "Lookup Result"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Lookup table data (columns B and C serve as the lookup range)
    # Column A: Product codes used as lookup keys
    # Column B: Category names
    # Column C: Unit prices (the lookup target values)
    data = [
        ["WDG-101", "Electronics", 249.99],
        ["WDG-102", "Electronics", 189.50],
        ["FRN-201", "Furniture", 575.00],
        ["WDG-103", "Electronics", 329.00],
        ["FRN-202", "Furniture", 1250.00],
        ["OFS-301", "Office Supplies", 42.75],
        ["OFS-302", "Office Supplies", 18.90],
        ["FRN-203", "Furniture", 899.00],
        ["WDG-104", "Electronics", 79.99],
        ["OFS-303", "Office Supplies", 64.50],
        ["FRN-204", "Furniture", 450.00],
        ["WDG-105", "Electronics", 549.00],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 3:
                cell.number_format = '$#,##0.00'

    # Column D: Lookup Result
    # Rows 2-4 and 6-13: leave empty (no formula)
    # Row 5 (D5): the broken formula with VLOKUP typo
    cell_d5 = ws.cell(row=5, column=4, value='=VLOKUP(A5,B:C,2,0)')
    cell_d5.border = thin_border
    cell_d5.number_format = '$#,##0.00'

    # Set column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16

    # Freeze header row
    ws.freeze_panes = "A2"

    # --- Sheet: Inventory ---
    ws2 = wb.create_sheet("Inventory")
    inv_headers = ["Product Code", "Warehouse", "Quantity", "Reorder Level", "Last Restocked"]
    for col, h in enumerate(inv_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    inv_data = [
        ["WDG-101", "East", 145, 50, "2025-11-20"],
        ["WDG-102", "East", 88, 30, "2025-10-15"],
        ["FRN-201", "West", 23, 10, "2025-09-05"],
        ["WDG-103", "East", 210, 75, "2025-12-01"],
        ["FRN-202", "Central", 7, 5, "2025-08-22"],
        ["OFS-301", "East", 520, 200, "2025-11-30"],
        ["OFS-302", "West", 340, 150, "2025-10-28"],
        ["FRN-203", "Central", 15, 8, "2025-07-14"],
        ["WDG-104", "East", 430, 100, "2025-12-10"],
        ["OFS-303", "West", 190, 80, "2025-11-05"],
        ["FRN-204", "Central", 31, 12, "2025-09-18"],
        ["WDG-105", "East", 67, 25, "2025-11-22"],
    ]

    for r, row_data in enumerate(inv_data, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
