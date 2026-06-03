"""
Initial Setup: Format SSN numbers with custom number format
Task ID: calc_lf_070
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_070'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: HR ---
    ws = wb.active
    ws.title = 'HR'

    # Headers
    ws['A1'] = 'Employee'
    ws['B1'] = 'SSN'

    # Style headers
    for cell in [ws['A1'], ws['B1']]:
        cell.font = Font(bold=True, size=11)

    # Data rows - SSN stored as integers (no formatting applied)
    ws['A2'] = 'Alice'
    ws['B2'] = 123456789

    ws['A3'] = 'Bob'
    ws['B3'] = 987654321

    ws['A4'] = 'Carol'
    ws['B4'] = 12345678  # 8-digit number, needs leading zero with format

    # Set column widths for readability
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18

    # B2:B4 remain with default 'General' number format (NO custom formatting)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
