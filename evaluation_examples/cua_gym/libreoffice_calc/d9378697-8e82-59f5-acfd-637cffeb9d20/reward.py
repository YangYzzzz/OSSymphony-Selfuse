"""
Reward Script: Format SSN numbers as 'XXX-XX-XXXX' using format '000-00-0000'
Task ID: calc_lf_070
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.35): B2 number format is '000-00-0000'
  - Component 2 (0.35): B3 number format is '000-00-0000'
  - Component 3 (0.30): B4 number format is '000-00-0000'
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_070'
EXPECTED_FORMAT = '000-00-0000'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'HR' sheet must exist
    if 'HR' not in wb.sheetnames:
        print("CRITICAL: Sheet 'HR' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['HR']

    # Component 1: B2 has SSN format '000-00-0000' (0.35 points)
    try:
        fmt = ws['B2'].number_format
        if fmt == EXPECTED_FORMAT:
            print(f"PASS: Component 1 — B2 number_format is '{fmt}' (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 1 — B2 number_format expected '{EXPECTED_FORMAT}', found '{fmt}'")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: B3 has SSN format '000-00-0000' (0.35 points)
    try:
        fmt = ws['B3'].number_format
        if fmt == EXPECTED_FORMAT:
            print(f"PASS: Component 2 — B3 number_format is '{fmt}' (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 2 — B3 number_format expected '{EXPECTED_FORMAT}', found '{fmt}'")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: B4 has SSN format '000-00-0000' (0.30 points)
    try:
        fmt = ws['B4'].number_format
        if fmt == EXPECTED_FORMAT:
            print(f"PASS: Component 3 — B4 number_format is '{fmt}' (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 3 — B4 number_format expected '{EXPECTED_FORMAT}', found '{fmt}'")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
