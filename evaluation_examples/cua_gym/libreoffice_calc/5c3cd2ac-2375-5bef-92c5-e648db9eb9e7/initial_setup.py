"""
Initial Setup: Two-level row outline grouping in LibreOffice Calc
Task ID: calc_adv_group_outline_levels_056
Domain: libreoffice_calc

Creates a 'Hierarchical Report' sheet with a two-level row outline:
- Level 1 group: rows 2-40 (the outermost group)
- Level 2 groups: rows 2-10, 12-20, 22-30, 32-40
- Level 2 summary rows: rows 11, 21, 31, 41
- Level 1 summary row: row 41
- All rows visible (all groups expanded)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_adv_group_outline_levels_056'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Hierarchical Report'

    # ---- Headers (row 1) ----
    headers = ['Region', 'Product', 'Q1 Sales', 'Q2 Sales', 'Q3 Sales', 'Q4 Sales', 'Annual Total']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFFFF')
        cell.alignment = Alignment(horizontal='center')

    # ---- Group 1: North America — rows 2-10 (detail), row 11 (subtotal) ----
    group1_data = [
        ('North America', 'Software Licenses',   142500, 138200, 155300, 162100),
        ('North America', 'Hardware',              89300,  92100,  88700,  95200),
        ('North America', 'Cloud Services',       203400, 215600, 228900, 241300),
        ('North America', 'Professional Services', 67800,  71200,  74500,  78900),
        ('North America', 'Maintenance',           45600,  46800,  48200,  49700),
        ('North America', 'Training',              23400,  24100,  25300,  26800),
        ('North America', 'Consulting',            56700,  58900,  61200,  63500),
        ('North America', 'Support Contracts',     34500,  35800,  37100,  38600),
        ('North America', 'Custom Development',    78900,  81200,  83600,  86100),
    ]
    for i, (region, product, q1, q2, q3, q4) in enumerate(group1_data, 2):
        ws.cell(row=i, column=1, value=region)
        ws.cell(row=i, column=2, value=product)
        ws.cell(row=i, column=3, value=q1)
        ws.cell(row=i, column=4, value=q2)
        ws.cell(row=i, column=5, value=q3)
        ws.cell(row=i, column=6, value=q4)
        ws.cell(row=i, column=7, value=f'=C{i}+D{i}+E{i}+F{i}')
        ws.row_dimensions[i].outlineLevel = 2

    # Row 11: North America subtotal (level 2 summary)
    ws.cell(row=11, column=1, value='North America Subtotal')
    ws.cell(row=11, column=2, value='')
    ws.cell(row=11, column=3, value='=SUM(C2:C10)')
    ws.cell(row=11, column=4, value='=SUM(D2:D10)')
    ws.cell(row=11, column=5, value='=SUM(E2:E10)')
    ws.cell(row=11, column=6, value='=SUM(F2:F10)')
    ws.cell(row=11, column=7, value='=SUM(G2:G10)')
    ws.row_dimensions[11].outlineLevel = 1
    for col in range(1, 8):
        ws.cell(row=11, column=col).font = Font(bold=True)
        ws.cell(row=11, column=col).fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')

    # ---- Group 2: Europe — rows 12-20 (detail), row 21 (subtotal) ----
    group2_data = [
        ('Europe', 'Software Licenses',    98700,  102300, 107800, 113200),
        ('Europe', 'Hardware',              61400,   63800,  66100,  68500),
        ('Europe', 'Cloud Services',       145200,  152300, 159700, 167400),
        ('Europe', 'Professional Services', 51300,   53800,  56400,  59100),
        ('Europe', 'Maintenance',           33200,   34500,  35900,  37300),
        ('Europe', 'Training',              17800,   18600,  19500,  20500),
        ('Europe', 'Consulting',            42100,   44300,  46600,  48900),
        ('Europe', 'Support Contracts',     25600,   26800,  28100,  29500),
        ('Europe', 'Custom Development',    58300,   61200,  64300,  67500),
    ]
    for i, (region, product, q1, q2, q3, q4) in enumerate(group2_data, 12):
        ws.cell(row=i, column=1, value=region)
        ws.cell(row=i, column=2, value=product)
        ws.cell(row=i, column=3, value=q1)
        ws.cell(row=i, column=4, value=q2)
        ws.cell(row=i, column=5, value=q3)
        ws.cell(row=i, column=6, value=q4)
        ws.cell(row=i, column=7, value=f'=C{i}+D{i}+E{i}+F{i}')
        ws.row_dimensions[i].outlineLevel = 2

    # Row 21: Europe subtotal (level 2 summary)
    ws.cell(row=21, column=1, value='Europe Subtotal')
    ws.cell(row=21, column=2, value='')
    ws.cell(row=21, column=3, value='=SUM(C12:C20)')
    ws.cell(row=21, column=4, value='=SUM(D12:D20)')
    ws.cell(row=21, column=5, value='=SUM(E12:E20)')
    ws.cell(row=21, column=6, value='=SUM(F12:F20)')
    ws.cell(row=21, column=7, value='=SUM(G12:G20)')
    ws.row_dimensions[21].outlineLevel = 1
    for col in range(1, 8):
        ws.cell(row=21, column=col).font = Font(bold=True)
        ws.cell(row=21, column=col).fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')

    # ---- Group 3: Asia Pacific — rows 22-30 (detail), row 31 (subtotal) ----
    group3_data = [
        ('Asia Pacific', 'Software Licenses',    76300,   79800,  83500,  87400),
        ('Asia Pacific', 'Hardware',              48900,   51200,  53600,  56100),
        ('Asia Pacific', 'Cloud Services',       112600,  118300, 124300, 130500),
        ('Asia Pacific', 'Professional Services', 39700,   41900,  44200,  46600),
        ('Asia Pacific', 'Maintenance',           26100,   27400,  28800,  30300),
        ('Asia Pacific', 'Training',              13900,   14700,  15600,  16600),
        ('Asia Pacific', 'Consulting',            33400,   35300,  37300,  39400),
        ('Asia Pacific', 'Support Contracts',     20200,   21400,  22700,  24100),
        ('Asia Pacific', 'Custom Development',    46700,   49400,  52200,  55100),
    ]
    for i, (region, product, q1, q2, q3, q4) in enumerate(group3_data, 22):
        ws.cell(row=i, column=1, value=region)
        ws.cell(row=i, column=2, value=product)
        ws.cell(row=i, column=3, value=q1)
        ws.cell(row=i, column=4, value=q2)
        ws.cell(row=i, column=5, value=q3)
        ws.cell(row=i, column=6, value=q4)
        ws.cell(row=i, column=7, value=f'=C{i}+D{i}+E{i}+F{i}')
        ws.row_dimensions[i].outlineLevel = 2

    # Row 31: Asia Pacific subtotal (level 2 summary)
    ws.cell(row=31, column=1, value='Asia Pacific Subtotal')
    ws.cell(row=31, column=2, value='')
    ws.cell(row=31, column=3, value='=SUM(C22:C30)')
    ws.cell(row=31, column=4, value='=SUM(D22:D30)')
    ws.cell(row=31, column=5, value='=SUM(E22:E30)')
    ws.cell(row=31, column=6, value='=SUM(F22:F30)')
    ws.cell(row=31, column=7, value='=SUM(G22:G30)')
    ws.row_dimensions[31].outlineLevel = 1
    for col in range(1, 8):
        ws.cell(row=31, column=col).font = Font(bold=True)
        ws.cell(row=31, column=col).fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')

    # ---- Group 4: Latin America — rows 32-40 (detail), row 41 (grand total) ----
    group4_data = [
        ('Latin America', 'Software Licenses',    45100,   47300,  49700,  52300),
        ('Latin America', 'Hardware',              28700,   30100,  31600,  33200),
        ('Latin America', 'Cloud Services',        67200,   70600,  74200,  78000),
        ('Latin America', 'Professional Services', 23600,   24900,  26300,  27800),
        ('Latin America', 'Maintenance',           15800,   16600,  17500,  18500),
        ('Latin America', 'Training',               8400,    8900,   9400,  10000),
        ('Latin America', 'Consulting',            19800,   20900,  22100,  23400),
        ('Latin America', 'Support Contracts',     12100,   12800,  13600,  14400),
        ('Latin America', 'Custom Development',    27800,   29400,  31100,  32900),
    ]
    for i, (region, product, q1, q2, q3, q4) in enumerate(group4_data, 32):
        ws.cell(row=i, column=1, value=region)
        ws.cell(row=i, column=2, value=product)
        ws.cell(row=i, column=3, value=q1)
        ws.cell(row=i, column=4, value=q2)
        ws.cell(row=i, column=5, value=q3)
        ws.cell(row=i, column=6, value=q4)
        ws.cell(row=i, column=7, value=f'=C{i}+D{i}+E{i}+F{i}')
        ws.row_dimensions[i].outlineLevel = 2

    # Row 41: Grand Total (level 1 summary — this is the ONLY row visible after level 1 collapse)
    ws.cell(row=41, column=1, value='Grand Total')
    ws.cell(row=41, column=2, value='')
    ws.cell(row=41, column=3, value='=SUM(C11,C21,C31)+SUM(C32:C40)')
    ws.cell(row=41, column=4, value='=SUM(D11,D21,D31)+SUM(D32:D40)')
    ws.cell(row=41, column=5, value='=SUM(E11,E21,E31)+SUM(E32:E40)')
    ws.cell(row=41, column=6, value='=SUM(F11,F21,F31)+SUM(F32:F40)')
    ws.cell(row=41, column=7, value='=SUM(G11,G21,G31)+SUM(G32:G40)')
    ws.row_dimensions[41].outlineLevel = 0
    for col in range(1, 8):
        ws.cell(row=41, column=col).font = Font(bold=True, size=12)
        ws.cell(row=41, column=col).fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        ws.cell(row=41, column=col).font = Font(bold=True, size=12, color='FFFFFFFF')

    # ---- Column widths ----
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 26
    for col_letter in ['C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].width = 14

    # ---- Freeze header row ----
    ws.freeze_panes = 'A2'

    # ---- All rows explicitly visible (no rows hidden) ----
    for row in range(1, 42):
        ws.row_dimensions[row].hidden = False

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet: Hierarchical Report with {ws.max_row} rows')
    print(f'Level 2 outline rows: 2-10, 12-20, 22-30, 32-40')
    print(f'Level 1 outline rows: 11, 21, 31 (subtotals), outline level 0 row: 41 (grand total)')
    print(f'All rows visible (expanded state)')


create_initial()
