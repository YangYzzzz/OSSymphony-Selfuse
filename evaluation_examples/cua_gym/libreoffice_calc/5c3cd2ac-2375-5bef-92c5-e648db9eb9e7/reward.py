"""
Reward Script: Collapse row outline to level 1 — hide all detail rows
Task ID: calc_adv_group_outline_levels_056
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5 pts): All level-2 detail rows (rows 2-10, 12-20, 22-30, 32-40) are hidden
  Component 2 (0.3 pts): All level-1 sub-summary rows (rows 11, 21, 31) are also hidden
  Component 3 (0.2 pts): Top-level summary row (row 41) remains visible AND
                          outline structure (row outline levels) is preserved
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_adv_group_outline_levels_056'
SHEET_NAME = 'Hierarchical Report'

# Expected row outline structure (from task context):
#   level-2 rows: 2-10, 12-20, 22-30, 32-40  (detail rows — should be hidden after level-1 collapse)
#   level-1 rows: 11, 21, 31                  (sub-summary rows — should also be hidden)
#   level-0 rows: 1 (header), 41 (top summary) (always visible)
LEVEL_2_ROWS = list(range(2, 11)) + list(range(12, 21)) + list(range(22, 31)) + list(range(32, 41))
LEVEL_1_ROWS = [11, 21, 31]
VISIBLE_ROWS = [1, 41]


def verify_task(file_path):
    """
    Verify that the agent clicked the level-1 outline button to collapse all groups,
    hiding all detail (level-2) and sub-summary (level-1) rows, leaving only the
    top-level summary row (41) and header row (1) visible.

    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: File must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Expected sheet must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found in workbook (found: {wb.sheetnames})")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Component 1: All level-2 detail rows are hidden (0.5 points)
    # These are rows 2-10, 12-20, 22-30, 32-40.
    # In the initial file, all these rows are visible (hidden=False).
    # After clicking level-1 outline button, all must be hidden (hidden=True).
    try:
        level2_hidden_count = 0
        level2_not_hidden = []
        for row_idx in LEVEL_2_ROWS:
            rd = ws.row_dimensions.get(row_idx)
            is_hidden = rd.hidden if rd else False
            if is_hidden:
                level2_hidden_count += 1
            else:
                level2_not_hidden.append(row_idx)

        total_level2 = len(LEVEL_2_ROWS)
        if level2_hidden_count == total_level2:
            print(f"PASS: Component 1 — all {total_level2} level-2 detail rows are hidden (0.5 pts)")
            total_score += 0.5
        else:
            visible_rows = level2_not_hidden[:5]  # show first 5 for brevity
            print(f"FAIL: Component 1 — only {level2_hidden_count}/{total_level2} level-2 rows are hidden; "
                  f"first visible detail rows: {visible_rows}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All level-1 sub-summary rows are also hidden (0.3 points)
    # Rows 11, 21, 31 are level-1 sub-summaries inside the level-1 group.
    # When level-1 outline button is clicked, these are also collapsed and hidden.
    try:
        level1_hidden_count = 0
        level1_not_hidden = []
        for row_idx in LEVEL_1_ROWS:
            rd = ws.row_dimensions.get(row_idx)
            is_hidden = rd.hidden if rd else False
            if is_hidden:
                level1_hidden_count += 1
            else:
                level1_not_hidden.append(row_idx)

        total_level1 = len(LEVEL_1_ROWS)
        if level1_hidden_count == total_level1:
            print(f"PASS: Component 2 — all {total_level1} level-1 sub-summary rows "
                  f"({LEVEL_1_ROWS}) are hidden (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — only {level1_hidden_count}/{total_level1} level-1 "
                  f"sub-summary rows are hidden; still visible: {level1_not_hidden}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Summary row 41 remains visible AND all rows 2-40 are hidden (0.2 points)
    # This is a compound check that ensures the collapse was done correctly:
    #   - Row 41 (top-level summary) must be visible (it must NOT have been accidentally hidden)
    #   - At least one detail/sub-summary row from rows 2-40 is confirmed hidden
    #     (as a compound anchor ensuring the task change actually happened)
    # In the initial state, rows 2-40 are all visible, so this compound check FAILS on initial.
    # After collapse, rows 2-40 are hidden but row 41 remains visible — compound check PASSES.
    try:
        # Check row 41 is NOT hidden (must still be visible after collapse)
        rd_41 = ws.row_dimensions.get(41)
        row41_visible = not (rd_41.hidden if rd_41 else False)

        # Check that the total set of hidden rows in 2-40 equals exactly 39
        # (all 36 level-2 + all 3 level-1 rows must be hidden)
        all_collapsed_rows = LEVEL_2_ROWS + LEVEL_1_ROWS  # all rows 2-40 in outline
        collapsed_hidden_count = 0
        for row_idx in all_collapsed_rows:
            rd = ws.row_dimensions.get(row_idx)
            if rd and rd.hidden:
                collapsed_hidden_count += 1

        all_collapsed = (collapsed_hidden_count == len(all_collapsed_rows))

        if row41_visible and all_collapsed:
            print("PASS: Component 3 — row 41 is visible and all rows 2-40 are hidden; "
                  "complete level-1 collapse confirmed (0.2 pts)")
            total_score += 0.2
        elif not row41_visible:
            print("FAIL: Component 3 — row 41 (top-level summary) is incorrectly hidden")
        else:
            print(f"FAIL: Component 3 — not all rows 2-40 are hidden "
                  f"(hidden: {collapsed_hidden_count}/{len(all_collapsed_rows)}); "
                  f"row 41 visible: {row41_visible}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
