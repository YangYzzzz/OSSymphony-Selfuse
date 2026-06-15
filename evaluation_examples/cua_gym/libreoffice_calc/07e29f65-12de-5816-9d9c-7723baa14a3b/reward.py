"""
Reward Script: Create a pivot table with year-over-year comparison
Task ID: calc_pivot_042
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20): Pivot sheet exists (not in initial)
  Component 2 (0.15): Column headers are Category, 2023, 2024
  Component 3 (0.15): Row labels contain Products, Services, Subscriptions
  Component 4 (0.30): Revenue values match expected (6 cells, 0.05 each)
  Component 5 (0.20): Grand Total row with correct sums
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_042'

# Expected values from task context
EXPECTED_VALUES = {
    'Services': {'2023': 95000, '2024': 112000},
    'Products': {'2023': 145000, '2024': 168000},
    'Subscriptions': {'2023': 60000, '2024': 78000},
}
EXPECTED_GRAND_TOTAL = {'2023': 300000, '2024': 358000}

# Tolerance for value matching (allow rounding differences)
TOLERANCE = 500


def find_pivot_sheet(wb):
    """Find a sheet that looks like a pivot table (not 'YoYData')."""
    for name in wb.sheetnames:
        if name.lower() != 'yoydata':
            ws = wb[name]
            # Check if it has at least a few rows and looks pivot-like
            if ws.max_row >= 4 and ws.max_column >= 3:
                return ws
    return None


def normalize(val):
    """Normalize a cell value to string for comparison."""
    if val is None:
        return ''
    return str(val).strip()


def numeric_val(val):
    """Extract numeric value from a cell."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(',', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Pivot sheet exists (0.20 points)
    # This FAILS on initial (only 'YoYData') and PASSES on golden
    try:
        pivot_ws = find_pivot_sheet(wb)
        if pivot_ws is not None:
            print(f"PASS: Component 1 — Pivot sheet '{pivot_ws.title}' exists (0.20 pts)")
            total_score += 0.20
        else:
            print("FAIL: Component 1 — No pivot sheet found (only 'YoYData' present)")
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Read the pivot sheet structure
    # Find header row and build column map
    ws = pivot_ws

    # Component 2: Column headers include year columns for 2023 and 2024 (0.15 points)
    try:
        # Scan first 5 rows for header row containing '2023' and '2024'
        header_row = None
        cat_col = None
        col_2023 = None
        col_2024 = None

        for r in range(1, min(6, ws.max_row + 1)):
            row_vals = {}
            for c in range(1, ws.max_column + 1):
                val = normalize(ws.cell(row=r, column=c).value)
                row_vals[c] = val

            # Check if this row has year identifiers
            has_2023 = any('2023' in v for v in row_vals.values())
            has_2024 = any('2024' in v for v in row_vals.values())
            has_cat = any('categ' in v.lower() for v in row_vals.values())

            if has_2023 and has_2024:
                header_row = r
                for c, v in row_vals.items():
                    if 'categ' in v.lower():
                        cat_col = c
                    if '2023' in v:
                        col_2023 = c
                    if '2024' in v:
                        col_2024 = c
                # If no explicit category column, assume first column
                if cat_col is None:
                    cat_col = 1
                break

        if header_row is not None and col_2023 is not None and col_2024 is not None:
            print(f"PASS: Component 2 — Headers found at row {header_row}: "
                  f"Category(col {cat_col}), 2023(col {col_2023}), 2024(col {col_2024}) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 — Could not find year-based column headers (2023 & 2024)")
            # Can't proceed without headers
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Component 3: Row labels contain all 3 categories (0.15 points)
    try:
        found_categories = {}
        for r in range(header_row + 1, ws.max_row + 1):
            cat_val = normalize(ws.cell(row=r, column=cat_col).value).lower()
            for expected_cat in EXPECTED_VALUES:
                if expected_cat.lower() in cat_val:
                    found_categories[expected_cat] = r

        if len(found_categories) >= 3:
            print(f"PASS: Component 3 — All 3 categories found: {list(found_categories.keys())} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 — Found {len(found_categories)}/3 categories: {list(found_categories.keys())}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Revenue values match expected for each category-year (0.30 points)
    # 6 cells total, 0.05 each
    try:
        value_score = 0.0
        for cat, row_num in found_categories.items():
            expected = EXPECTED_VALUES[cat]

            # Check 2023
            actual_2023 = numeric_val(ws.cell(row=row_num, column=col_2023).value)
            if actual_2023 is not None and abs(actual_2023 - expected['2023']) <= TOLERANCE:
                print(f"PASS: Component 4 — {cat}/2023 = {actual_2023} (expected ~{expected['2023']}) (0.05 pts)")
                value_score += 0.05
            else:
                print(f"FAIL: Component 4 — {cat}/2023 = {actual_2023} (expected ~{expected['2023']})")

            # Check 2024
            actual_2024 = numeric_val(ws.cell(row=row_num, column=col_2024).value)
            if actual_2024 is not None and abs(actual_2024 - expected['2024']) <= TOLERANCE:
                print(f"PASS: Component 4 — {cat}/2024 = {actual_2024} (expected ~{expected['2024']}) (0.05 pts)")
                value_score += 0.05
            else:
                print(f"FAIL: Component 4 — {cat}/2024 = {actual_2024} (expected ~{expected['2024']})")

        total_score += value_score
        print(f"Component 4 subtotal: {value_score}/0.30")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Grand Total row with correct sums (0.20 points)
    try:
        grand_total_row = None
        for r in range(header_row + 1, ws.max_row + 1):
            cat_val = normalize(ws.cell(row=r, column=cat_col).value).lower()
            if 'total' in cat_val or 'grand' in cat_val:
                grand_total_row = r
                break

        if grand_total_row is not None:
            gt_2023 = numeric_val(ws.cell(row=grand_total_row, column=col_2023).value)
            gt_2024 = numeric_val(ws.cell(row=grand_total_row, column=col_2024).value)

            gt_pass = 0
            if gt_2023 is not None and abs(gt_2023 - EXPECTED_GRAND_TOTAL['2023']) <= TOLERANCE:
                gt_pass += 1
                print(f"PASS: Component 5a — Grand Total 2023 = {gt_2023} (expected ~{EXPECTED_GRAND_TOTAL['2023']})")
            else:
                print(f"FAIL: Component 5a — Grand Total 2023 = {gt_2023} (expected ~{EXPECTED_GRAND_TOTAL['2023']})")

            if gt_2024 is not None and abs(gt_2024 - EXPECTED_GRAND_TOTAL['2024']) <= TOLERANCE:
                gt_pass += 1
                print(f"PASS: Component 5b — Grand Total 2024 = {gt_2024} (expected ~{EXPECTED_GRAND_TOTAL['2024']})")
            else:
                print(f"FAIL: Component 5b — Grand Total 2024 = {gt_2024} (expected ~{EXPECTED_GRAND_TOTAL['2024']})")

            if gt_pass == 2:
                total_score += 0.20
                print(f"PASS: Component 5 — Grand Total row correct (0.20 pts)")
            elif gt_pass == 1:
                total_score += 0.10
                print(f"PARTIAL: Component 5 — 1/2 grand totals correct (0.10 pts)")
            else:
                print(f"FAIL: Component 5 — Grand Total row values incorrect")
        else:
            print(f"FAIL: Component 5 — No Grand Total row found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
