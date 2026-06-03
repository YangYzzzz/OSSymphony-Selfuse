"""
Initial Setup: Create YoY revenue data for pivot table task
Task ID: calc_pivot_042
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_042'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def random_dates_in_year(year, count):
    """Generate count random dates within a given year."""
    start = date(year, 1, 1)
    end = date(year, 12, 31)
    days_range = (end - start).days
    dates = []
    for _ in range(count):
        d = start + timedelta(days=random.randint(0, days_range))
        dates.append(d)
    return dates


def distribute_revenue(total, count):
    """Distribute total revenue across count transactions realistically."""
    # Generate random weights and scale to match total
    raw = [random.uniform(50, 800) for _ in range(count)]
    raw_sum = sum(raw)
    scaled = [round(r / raw_sum * total, 2) for r in raw]
    # Fix rounding error on last element
    diff = round(total - sum(scaled), 2)
    scaled[-1] = round(scaled[-1] + diff, 2)
    return scaled


def create_initial():
    random.seed(42)  # reproducible

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'YoYData'

    # Headers
    headers = ['TxnID', 'TxnDate', 'Category', 'Revenue']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.font = Font(bold=True, size=11, color="FFFFFF")

    # Revenue targets per category/year
    targets = {
        ('Services', 2023): 95000,
        ('Services', 2024): 112000,
        ('Products', 2023): 145000,
        ('Products', 2024): 168000,
        ('Subscriptions', 2023): 60000,
        ('Subscriptions', 2024): 78000,
    }

    # 600 rows total, 300 per year, 100 per category per year
    categories = ['Services', 'Products', 'Subscriptions']
    rows_per_cat_per_year = 100

    all_rows = []

    for year in [2023, 2024]:
        dates = random_dates_in_year(year, 300)
        date_idx = 0
        for cat in categories:
            total = targets[(cat, year)]
            revenues = distribute_revenue(total, rows_per_cat_per_year)
            for i in range(rows_per_cat_per_year):
                all_rows.append((dates[date_idx], cat, revenues[i]))
                date_idx += 1

    # Shuffle to make it look realistic (not grouped by category)
    random.shuffle(all_rows)

    # Write data
    for r, (txn_date, cat, rev) in enumerate(all_rows, 2):
        ws.cell(row=r, column=1, value=r - 1)  # TxnID 1-600
        ws.cell(row=r, column=2, value=txn_date)
        ws.cell(row=r, column=2).number_format = 'MM/DD/YYYY'
        ws.cell(row=r, column=3, value=cat)
        ws.cell(row=r, column=4, value=rev)
        ws.cell(row=r, column=4).number_format = '#,##0.00'

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Verify totals
    from collections import defaultdict
    totals = defaultdict(float)
    for txn_date, cat, rev in all_rows:
        totals[(cat, txn_date.year)] += rev
    for key, val in sorted(totals.items()):
        print(f'  {key}: {val:.2f}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
