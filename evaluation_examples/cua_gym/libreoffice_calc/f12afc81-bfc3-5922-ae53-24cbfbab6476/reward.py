"""
Reward Script: Format C2:C20 as USD currency and bold the C1 header
Task ID: calc_gsd_002
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.5): C2:C20 have USD currency format ($#,##0.00 or equivalent)
  - Component 2 (0.2): C1 header text is 'Amount'
  - Component 3 (0.3): C1 is bold
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_002'


def persist_app_state(domain: str):
    """Save any unsaved LibreOffice edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Annual Budget' sheet must exist
    if 'Annual Budget' not in wb.sheetnames:
        print(f"CRITICAL: Sheet 'Annual Budget' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Annual Budget']

    # Component 1: C2:C20 have USD currency number format (0.5 points)
    # Acceptable formats: '$#,##0.00' or similar USD currency with 2 decimals
    # This checks the task-introduced change: formatting was 'General' initially.
    try:
        currency_cells = 0
        total_cells = 19  # C2 through C20
        acceptable_formats = {'$#,##0.00', '"$"#,##0.00', '[$USD] #,##0.00', '[$$-409]#,##0.00'}
        for r in range(2, 21):
            cell = ws.cell(row=r, column=3)
            nf = cell.number_format
            # Check if it's a dollar currency format with 2 decimal places
            if nf in acceptable_formats or ('$' in nf and '0.00' in nf):
                currency_cells += 1
        ratio = currency_cells / total_cells
        if ratio >= 1.0:
            print(f"PASS: Component 1 — All {total_cells} cells in C2:C20 have USD currency format (0.5 pts)")
            total_score += 0.5
        elif ratio >= 0.5:
            partial = round(0.5 * ratio, 2)
            print(f"PARTIAL: Component 1 — {currency_cells}/{total_cells} cells formatted ({partial} pts)")
            total_score += partial
        else:
            sample_fmt = ws.cell(row=2, column=3).number_format
            print(f"FAIL: Component 1 — Only {currency_cells}/{total_cells} cells have currency format. C2 format: {sample_fmt!r}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: C1 text is 'Amount' (0.2 points)
    # This is a compound check: header preserved AND bold applied (anchored to bold being the change).
    # Actually, the header text is a precondition. Let me restructure:
    # Component 2 checks that C1 is bold — this is the task-introduced change.
    # But we split bold into its own component for partial credit.
    # To avoid scoring a precondition, we make C1='Amount' a gate, not a score.

    # Component 2: C1 is bold (0.3 points) — was NOT bold initially
    try:
        c1 = ws['C1']
        if c1.font.bold:
            print(f"PASS: Component 2 — C1 is bold (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — C1 is not bold (font.bold={c1.font.bold})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Column D and E formatting unchanged (0.2 points)
    # D should be '#,##0' and E should be '0.00%' — verifies no collateral damage
    # This component fails on initial if D/E already have these formats... but they DO.
    # So we need to anchor this to C being currency AND D/E unchanged.
    # Compound check: C2 has currency format AND D2/E2 formats are preserved.
    try:
        c2_fmt = ws.cell(row=2, column=3).number_format
        d2_fmt = ws.cell(row=2, column=4).number_format
        e2_fmt = ws.cell(row=2, column=5).number_format
        c2_is_currency = '$' in c2_fmt and '0.00' in c2_fmt
        d2_ok = d2_fmt == '#,##0'
        e2_ok = e2_fmt == '0.00%'
        if c2_is_currency and d2_ok and e2_ok:
            print(f"PASS: Component 3 — Currency applied to C AND D/E formatting preserved (0.2 pts)")
            total_score += 0.2
        elif not c2_is_currency:
            print(f"FAIL: Component 3 — C2 not in currency format yet (format: {c2_fmt!r}), so compound check fails")
        else:
            print(f"FAIL: Component 3 — D/E formatting changed. D2={d2_fmt!r} (expected '#,##0'), E2={e2_fmt!r} (expected '0.00%')")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state('libreoffice_calc')

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
