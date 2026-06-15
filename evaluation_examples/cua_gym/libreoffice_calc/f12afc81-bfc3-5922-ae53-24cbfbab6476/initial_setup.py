"""
Initial Setup: Format cells C2:C20 as currency and bold C1 header
Task ID: calc_gsd_002
Domain: libreoffice_calc

Creates a budget tracker spreadsheet with 19 department budget rows.
Column C has plain numeric amounts (no currency formatting).
Cell C1 "Amount" is NOT bold. Column D has Variance formulas, Column E has % Change formulas.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_002'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Annual Budget"

    # --- Headers (Row 1) ---
    headers = ["Category", "Description", "Amount", "Variance", "% Change"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
        ws.cell(row=1, column=col).font = Font(name="Calibri", size=11)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

    # --- Column widths ---
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14

    # --- Data rows 2-20 (19 rows of department budget entries) ---
    data = [
        # (Category, Description, Amount, Planned Amount for Variance calc)
        ("Engineering", "Software development team salaries", 128500, 125000),
        ("Engineering", "Cloud infrastructure costs", 45000, 42000),
        ("Engineering", "Developer tools and licenses", 18700, 20000),
        ("Marketing", "Digital advertising campaigns", 67200, 70000),
        ("Marketing", "Brand design and collateral", 23400, 25000),
        ("Marketing", "Trade show exhibitions", 41500, 40000),
        ("Sales", "Sales team compensation", 95800, 92000),
        ("Sales", "CRM platform subscription", 12600, 12000),
        ("Sales", "Client entertainment expenses", 8900, 10000),
        ("Human Resources", "Recruitment and hiring costs", 34200, 35000),
        ("Human Resources", "Employee training programs", 27800, 28000),
        ("Human Resources", "Benefits administration", 52100, 50000),
        ("Finance", "Accounting software suite", 15300, 15000),
        ("Finance", "External audit fees", 38700, 40000),
        ("Finance", "Financial consulting services", 21400, 22000),
        ("Operations", "Office lease and utilities", 76500, 75000),
        ("Operations", "Equipment maintenance", 9200, 10000),
        ("Operations", "Supply chain logistics", 63800, 60000),
        ("Executive", "Executive compensation package", 145000, 140000),
    ]

    for r, (category, description, amount, planned) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=category)
        ws.cell(row=r, column=2, value=description)
        ws.cell(row=r, column=3, value=amount)  # Plain number, NO formatting
        # Column D: Variance = Amount - Planned (as formula referencing a hidden helper)
        # We store the variance as a simple formula: =C{r}-{planned}
        ws.cell(row=r, column=4, value=f'=C{r}-{planned}')
        ws.cell(row=r, column=4).number_format = '#,##0'
        # Column E: % Change = Variance / Planned
        ws.cell(row=r, column=5, value=f'=D{r}/{planned}')
        ws.cell(row=r, column=5).number_format = '0.00%'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
