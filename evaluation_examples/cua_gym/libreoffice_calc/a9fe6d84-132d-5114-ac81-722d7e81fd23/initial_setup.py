"""
Initial Setup: Reorder sheets in the workbook to match alphabetical order
Task ID: calc_gsi_079
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_079'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    def write_sheet(ws, headers, data):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        for r, row_data in enumerate(data, 2):
            for c, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = thin_border
                if isinstance(val, (int, float)) and headers[c - 1] in ("Budget ($)", "Salary ($)", "Annual Cost ($)", "Hourly Rate ($)", "Monthly Cost ($)", "Retainer ($)"):
                    cell.number_format = '$#,##0'

    # --- Sheet 1: Marketing (active sheet) ---
    ws_marketing = wb.active
    ws_marketing.title = "Marketing"
    marketing_headers = ["Campaign", "Channel", "Budget ($)", "Leads", "Start Date", "Status"]
    marketing_data = [
        ["Spring Launch", "Social Media", 25000, 1420, "2025-03-01", "Active"],
        ["Email Blast Q1", "Email", 8500, 890, "2025-01-15", "Completed"],
        ["Brand Refresh", "Print", 42000, 310, "2025-02-20", "Active"],
        ["Webinar Series", "Online Events", 12000, 675, "2025-04-05", "Planned"],
        ["PPC Campaign", "Google Ads", 35000, 2100, "2025-01-10", "Active"],
        ["Trade Show", "Events", 55000, 450, "2025-05-12", "Planned"],
        ["Influencer Collab", "Social Media", 18000, 1230, "2025-03-18", "Active"],
        ["Newsletter Redesign", "Email", 5000, 560, "2025-02-01", "Completed"],
        ["Video Series", "YouTube", 30000, 890, "2025-04-22", "Planned"],
        ["Podcast Sponsor", "Audio", 15000, 340, "2025-03-30", "Active"],
        ["SEO Overhaul", "Organic", 20000, 1680, "2025-01-05", "Active"],
        ["Holiday Promo", "Multi-channel", 48000, 3200, "2025-11-15", "Planned"],
    ]
    write_sheet(ws_marketing, marketing_headers, marketing_data)

    # --- Sheet 2: IT ---
    ws_it = wb.create_sheet("IT")
    it_headers = ["Asset ID", "Device Type", "Assigned To", "Annual Cost ($)", "Purchase Date", "Warranty Exp"]
    it_data = [
        ["IT-1001", "Laptop", "Sarah Chen", 2400, "2024-06-15", "2027-06-15"],
        ["IT-1002", "Desktop", "Marcus Johnson", 1800, "2024-03-10", "2027-03-10"],
        ["IT-1003", "Monitor", "Priya Patel", 650, "2024-09-22", "2026-09-22"],
        ["IT-1004", "Laptop", "James Wilson", 2400, "2025-01-08", "2028-01-08"],
        ["IT-1005", "Server Rack", "IT Dept", 15000, "2023-11-20", "2028-11-20"],
        ["IT-1006", "Printer", "3rd Floor", 3200, "2024-07-01", "2026-07-01"],
        ["IT-1007", "Laptop", "Emily Rodriguez", 2800, "2025-02-14", "2028-02-14"],
        ["IT-1008", "Tablet", "David Kim", 900, "2024-12-05", "2026-12-05"],
        ["IT-1009", "Network Switch", "Server Room", 4500, "2024-01-30", "2029-01-30"],
        ["IT-1010", "Laptop", "Lisa Thompson", 2400, "2024-08-18", "2027-08-18"],
        ["IT-1011", "Webcam", "Conference Rm A", 250, "2025-03-01", "2027-03-01"],
    ]
    write_sheet(ws_it, it_headers, it_data)

    # --- Sheet 3: Finance ---
    ws_finance = wb.create_sheet("Finance")
    finance_headers = ["Account", "Category", "Q1 Actual ($)", "Q2 Actual ($)", "Q3 Forecast ($)", "Variance (%)"]
    finance_data = [
        ["4100", "Product Revenue", 1250000, 1380000, 1420000, 4.2],
        ["4200", "Service Revenue", 340000, 365000, 380000, 2.8],
        ["5100", "COGS - Materials", 425000, 448000, 460000, -1.5],
        ["5200", "COGS - Labor", 310000, 322000, 330000, -0.8],
        ["6100", "Marketing Expense", 85000, 92000, 88000, 3.1],
        ["6200", "R&D Expense", 210000, 225000, 240000, -2.4],
        ["6300", "Admin Expense", 145000, 148000, 150000, 0.5],
        ["6400", "Facilities", 95000, 95000, 97000, 0.0],
        ["6500", "Travel & Ent.", 38000, 42000, 35000, 5.2],
        ["6600", "Insurance", 62000, 62000, 64000, 0.0],
        ["7100", "Interest Income", 12000, 13500, 14000, 1.8],
        ["7200", "Other Income", 8000, 9200, 8500, 3.0],
    ]
    write_sheet(ws_finance, finance_headers, finance_data)

    # --- Sheet 4: HR ---
    ws_hr = wb.create_sheet("HR")
    hr_headers = ["Employee ID", "Name", "Department", "Salary ($)", "Hire Date", "Performance"]
    hr_data = [
        ["EMP-001", "Sarah Chen", "Engineering", 125000, "2021-03-15", "Exceeds"],
        ["EMP-002", "Marcus Johnson", "Marketing", 92000, "2022-06-01", "Meets"],
        ["EMP-003", "Priya Patel", "Finance", 108000, "2020-09-10", "Exceeds"],
        ["EMP-004", "James Wilson", "Operations", 88000, "2023-01-20", "Meets"],
        ["EMP-005", "Emily Rodriguez", "Engineering", 115000, "2022-02-14", "Exceeds"],
        ["EMP-006", "David Kim", "IT", 98000, "2021-07-08", "Meets"],
        ["EMP-007", "Lisa Thompson", "HR", 95000, "2020-11-30", "Exceeds"],
        ["EMP-008", "Robert Garcia", "Legal", 135000, "2019-05-22", "Meets"],
        ["EMP-009", "Amanda Foster", "Finance", 102000, "2022-08-15", "Meets"],
        ["EMP-010", "Kevin Wright", "Operations", 85000, "2023-04-01", "Developing"],
        ["EMP-011", "Michelle Lee", "Marketing", 78000, "2024-01-10", "Meets"],
        ["EMP-012", "Daniel Brown", "IT", 105000, "2021-10-18", "Exceeds"],
    ]
    write_sheet(ws_hr, hr_headers, hr_data)

    # --- Sheet 5: Operations ---
    ws_ops = wb.create_sheet("Operations")
    ops_headers = ["Project", "Phase", "Monthly Cost ($)", "Team Size", "Deadline", "Priority"]
    ops_data = [
        ["Warehouse Expansion", "Construction", 180000, 24, "2025-09-30", "High"],
        ["Supply Chain Audit", "Analysis", 45000, 6, "2025-06-15", "Medium"],
        ["Fleet Upgrade", "Procurement", 320000, 8, "2025-12-01", "High"],
        ["Inventory System", "Implementation", 75000, 12, "2025-07-20", "High"],
        ["Safety Training", "Rollout", 28000, 4, "2025-05-10", "Medium"],
        ["Quality Control Rev", "Planning", 35000, 5, "2025-08-30", "Low"],
        ["Vendor Onboarding", "Execution", 22000, 3, "2025-04-25", "Medium"],
        ["Logistics Platform", "Testing", 95000, 15, "2025-10-15", "High"],
        ["Energy Efficiency", "Assessment", 18000, 3, "2025-06-30", "Low"],
        ["Compliance Update", "Documentation", 12000, 2, "2025-05-01", "Medium"],
    ]
    write_sheet(ws_ops, ops_headers, ops_data)

    # --- Sheet 6: Legal ---
    ws_legal = wb.create_sheet("Legal")
    legal_headers = ["Case ID", "Matter", "Client/Dept", "Retainer ($)", "Filed Date", "Status"]
    legal_data = [
        ["LGL-2025-001", "Patent Filing", "Engineering", 45000, "2025-01-20", "Pending"],
        ["LGL-2025-002", "Vendor Contract", "Operations", 12000, "2025-02-05", "Active"],
        ["LGL-2025-003", "Employment Dispute", "HR", 28000, "2025-01-30", "In Review"],
        ["LGL-2025-004", "Trademark Renewal", "Marketing", 8500, "2025-03-10", "Completed"],
        ["LGL-2025-005", "Lease Agreement", "Facilities", 18000, "2025-02-18", "Active"],
        ["LGL-2025-006", "Data Privacy Audit", "IT", 35000, "2025-04-01", "Planned"],
        ["LGL-2025-007", "M&A Due Diligence", "Finance", 120000, "2025-03-25", "Active"],
        ["LGL-2025-008", "IP Licensing", "Engineering", 22000, "2025-01-12", "Completed"],
        ["LGL-2025-009", "Regulatory Filing", "Compliance", 15000, "2025-02-28", "Pending"],
        ["LGL-2025-010", "NDA Review", "All Depts", 5000, "2025-03-15", "Active"],
    ]
    write_sheet(ws_legal, legal_headers, legal_data)

    # Sheet order is: Marketing, IT, Finance, HR, Operations, Legal (non-alphabetical)
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet order: {wb.sheetnames}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
