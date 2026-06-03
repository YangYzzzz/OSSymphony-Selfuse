"""
Reward Script: Reorder sheets alphabetically
Task ID: calc_gsi_079
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): Full sheet order matches alphabetical exactly
  Component 2 (0.3): Progressive partial credit per correctly-positioned sheet
  Component 3 (0.2): Data integrity - each sheet retains its expected header row
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_079'

EXPECTED_ORDER = ['Finance', 'HR', 'IT', 'Legal', 'Marketing', 'Operations']

# Expected first header cell (A1) for each sheet to verify data integrity
EXPECTED_HEADERS = {
    'Finance': 'Account',
    'HR': 'Employee ID',
    'IT': 'Asset ID',
    'Legal': 'Case ID',
    'Marketing': 'Campaign',
    'Operations': 'Project',
}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    actual_order = wb.sheetnames
    print(f"Actual sheet order: {actual_order}")
    print(f"Expected sheet order: {EXPECTED_ORDER}")

    # Precondition: all 6 expected sheets must exist
    missing = [s for s in EXPECTED_ORDER if s not in actual_order]
    if missing:
        print(f"CRITICAL: Missing sheets: {missing}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Full sheet order matches alphabetical exactly (0.5 points)
    # This is the primary task requirement - reorder all sheets alphabetically.
    # FAILS on initial (order is Marketing, IT, Finance, HR, Operations, Legal)
    # PASSES on golden (order is Finance, HR, IT, Legal, Marketing, Operations)
    try:
        if actual_order == EXPECTED_ORDER:
            print(f"PASS: Component 1 - Sheet order is exactly alphabetical (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 - Sheet order {actual_order} != expected {EXPECTED_ORDER}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Progressive partial credit - count correctly-positioned sheets (0.3 points)
    # Awards proportional credit based on how many sheets are in the right position.
    # Initial env has only Finance(idx2) and HR(idx3) at wrong positions... let me check:
    #   Initial: ['Marketing', 'IT', 'Finance', 'HR', 'Operations', 'Legal']
    #   Expected: ['Finance', 'HR', 'IT', 'Legal', 'Marketing', 'Operations']
    #   Position matches: none match -> 0/6 -> 0.0 pts on initial
    #   Golden: all 6 match -> 6/6 -> 0.3 pts on golden
    try:
        correct_positions = 0
        for i, expected_name in enumerate(EXPECTED_ORDER):
            if i < len(actual_order) and actual_order[i] == expected_name:
                correct_positions += 1
                print(f"  Position {i}: '{actual_order[i]}' correct")
            else:
                actual_at_pos = actual_order[i] if i < len(actual_order) else "N/A"
                print(f"  Position {i}: expected '{expected_name}', found '{actual_at_pos}'")

        # Only award points if MORE than 0 sheets are correctly positioned AND
        # not all are correct (that's already covered by Component 1).
        # But we also award when all correct, making it cumulative.
        # Key: initial has 0 correct positions -> 0 points. Golden has 6 -> 0.3 points.
        if correct_positions > 0:
            fraction = correct_positions / len(EXPECTED_ORDER)
            points = round(0.3 * fraction, 4)
            print(f"PASS: Component 2 - {correct_positions}/{len(EXPECTED_ORDER)} sheets in correct position ({points} pts)")
            total_score += points
        else:
            print(f"FAIL: Component 2 - No sheets in correct position (0 pts)")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Data integrity combined with correct ordering (0.2 points)
    # Verify that each sheet at its expected alphabetical position has the right header.
    # This ensures the sheet content was moved with the tab, not just renamed.
    # FAILS on initial because sheets are not in alphabetical order, so position-header
    # combinations won't match.
    try:
        integrity_ok = 0
        for i, expected_name in enumerate(EXPECTED_ORDER):
            if i < len(actual_order) and actual_order[i] == expected_name:
                ws = wb[expected_name]
                header_val = ws.cell(row=1, column=1).value
                expected_header = EXPECTED_HEADERS.get(expected_name)
                if header_val == expected_header:
                    integrity_ok += 1
                else:
                    print(f"  Data check '{expected_name}': expected header '{expected_header}', found '{header_val}'")

        if integrity_ok == len(EXPECTED_ORDER):
            print(f"PASS: Component 3 - All sheets at correct positions have correct data (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 - Only {integrity_ok}/{len(EXPECTED_ORDER)} sheets have correct data at correct position")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
