"""
Initial Setup: Create papers_by_topic.xlsx with academic papers listed by topic,
including ResNet (He et al., 2016) and papers that cite it.
Task ID: osworld_multi_apps_pdf_download_cite_004
Domain: libreoffice_calc (multi-app: calc + chrome)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_pdf_download_cite_004'
OUTPUT = f'{WORKDIR}/papers_by_topic.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Papers by Topic"

    # Header row
    headers = ["Topic", "Title", "URL"]
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align

    # Paper data: Topic, Title, URL
    # Row 1: ResNet — the foundational paper (first paper in list)
    papers = [
        (
            "Deep Learning / CNNs",
            "Deep Residual Learning for Image Recognition",
            "https://arxiv.org/pdf/1512.03385"
        ),
        (
            "Object Detection",
            "Faster R-CNN: Towards Real-Time Object Detection with Region Proposal Networks",
            "https://arxiv.org/pdf/1506.01497"
        ),
        (
            "Object Detection",
            "Feature Pyramid Networks for Object Detection",
            "https://arxiv.org/pdf/1612.03144"
        ),
        (
            "Object Detection",
            "You Only Look Once: Unified, Real-Time Object Detection",
            "https://arxiv.org/pdf/1506.02640"
        ),
        (
            "Object Detection",
            "SSD: Single Shot MultiBox Detector",
            "https://arxiv.org/pdf/1512.02325"
        ),
        (
            "Deep Learning / CNNs",
            "Very Deep Convolutional Networks for Large-Scale Image Recognition",
            "https://arxiv.org/pdf/1409.1556"
        ),
        (
            "Deep Learning / CNNs",
            "ImageNet Classification with Deep Convolutional Neural Networks",
            "https://proceedings.neurips.cc/paper_files/paper/2012/file/c399862d3b9d6b76c8436e924a68c45b-Paper.pdf"
        ),
        (
            "Semantic Segmentation",
            "Fully Convolutional Networks for Semantic Segmentation",
            "https://arxiv.org/pdf/1411.4038"
        ),
        (
            "Instance Segmentation",
            "Mask R-CNN",
            "https://arxiv.org/pdf/1703.06870"
        ),
        (
            "Image Recognition",
            "Densely Connected Convolutional Networks",
            "https://arxiv.org/pdf/1608.06993"
        ),
    ]

    for row_idx, (topic, title, url) in enumerate(papers, 2):
        ws.cell(row=row_idx, column=1, value=topic)
        ws.cell(row=row_idx, column=2, value=title)
        url_cell = ws.cell(row=row_idx, column=3, value=url)
        url_cell.font = Font(color="FF0563C1", underline="single")

    # Adjust column widths for readability
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 65

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f"Initial file created: {OUTPUT}")

    # Ensure /home/user/papers/ does NOT exist (task requires agent to create it)
    papers_dir = f"{WORKDIR}/papers"
    if os.path.exists(papers_dir):
        import shutil
        shutil.rmtree(papers_dir)
    print(f"Confirmed: {papers_dir} does not exist")

    # Ensure citing_paper.docx does NOT exist (task requires agent to create it)
    citing_docx = f"{WORKDIR}/citing_paper.docx"
    if os.path.exists(citing_docx):
        os.remove(citing_docx)
    print(f"Confirmed: {citing_docx} does not exist")

    # GUI-ready startup: open LibreOffice Calc with the spreadsheet, and Chrome
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    launch_gui('google-chrome', delay_sec=2.0)
    print("GUI_READY: launched LibreOffice Calc and Chrome with DISPLAY=:0")


create_initial()
