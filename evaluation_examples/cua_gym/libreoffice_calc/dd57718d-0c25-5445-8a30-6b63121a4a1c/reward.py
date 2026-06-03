"""
Reward Script: Create custom validation rule for cell D2
Task ID: calc_nrv_066
Domain: libreoffice_calc
Scoring:
  Component 1 (0.40): Custom data validation exists on D2
  Component 2 (0.35): Validation formula is =D2=B2+C2
  Component 3 (0.25): Error handling configured (showErrorMessage + error text)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_066'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Precondition: verify the basic data is intact (gate, not scored)
    if ws['B2'].value is None or ws['C2'].value is None:
        print("PRECONDITION FAIL: B2 or C2 missing — file may be corrupted")
        print("REWARD: 0.0")
        return 0.0

    # Find data validation targeting D2
    dv_on_d2 = None
    validations = ws.data_validations.dataValidation
    for dv in validations:
        # sqref can contain multiple ranges; check if D2 is covered
        sqref_str = str(dv.sqref)
        if 'D2' in sqref_str.upper().replace('$', ''):
            dv_on_d2 = dv
            break

    # Component 1: Custom data validation exists on D2 (0.40 points)
    try:
        if dv_on_d2 is not None and dv_on_d2.type == 'custom':
            print(f"PASS: Component 1 — Custom data validation found on D2 (type={dv_on_d2.type}) (0.40 pts)")
            total_score += 0.40
        elif dv_on_d2 is not None:
            print(f"FAIL: Component 1 — Data validation on D2 exists but type is '{dv_on_d2.type}', expected 'custom'")
        else:
            print(f"FAIL: Component 1 — No data validation found on cell D2 (found {len(validations)} total validations)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Validation formula is =D2=B2+C2 (0.35 points)
    try:
        if dv_on_d2 is not None and dv_on_d2.formula1 is not None:
            formula = str(dv_on_d2.formula1).upper().replace(" ", "").replace("$", "")
            expected = "=D2=B2+C2"
            if formula == expected.upper():
                print(f"PASS: Component 2 — Validation formula matches: {dv_on_d2.formula1} (0.35 pts)")
                total_score += 0.35
            else:
                print(f"FAIL: Component 2 — Expected formula '{expected}', found '{dv_on_d2.formula1}'")
        else:
            print(f"FAIL: Component 2 — No validation or no formula1 on D2")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Error handling configured (0.25 points)
    # showErrorMessage should be True and error text should be non-empty
    try:
        if dv_on_d2 is not None:
            has_error_msg = dv_on_d2.showErrorMessage is True or dv_on_d2.showErrorMessage is None
            has_error_text = dv_on_d2.error is not None and len(str(dv_on_d2.error).strip()) > 0
            if has_error_msg and has_error_text:
                print(f"PASS: Component 3 — Error handling configured (showErrorMessage={dv_on_d2.showErrorMessage}, error='{dv_on_d2.error}') (0.25 pts)")
                total_score += 0.25
            elif has_error_text:
                # Partial: has error text but showErrorMessage may be off
                print(f"PARTIAL: Component 3 — Error text present but showErrorMessage={dv_on_d2.showErrorMessage} (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 3 — showErrorMessage={dv_on_d2.showErrorMessage}, error='{dv_on_d2.error}'")
        else:
            print(f"FAIL: Component 3 — No validation on D2")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
