"""
Initial Setup: Create a spreadsheet with Item/Cost/Tax/Total headers and data, no validation.
Task ID: calc_nrv_066
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_066'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Headers
    ws["A1"] = "Item"
    ws["B1"] = "Cost"
    ws["C1"] = "Tax"
    ws["D1"] = "Total"

    # Data row - D2 intentionally left empty (task is to add validation there)
    ws["A2"] = "Widget"
    ws["B2"] = 100
    ws["C2"] = 8.50
    # D2 is empty - no value, no validation

    # Add more rows for realistic content
    ws["A3"] = "Gadget"
    ws["B3"] = 250
    ws["C3"] = 21.25

    ws["A4"] = "Sprocket"
    ws["B4"] = 75
    ws["C4"] = 6.38

    ws["A5"] = "Bracket"
    ws["B5"] = 40
    ws["C5"] = 3.40

    ws["A6"] = "Flange"
    ws["B6"] = 180
    ws["C6"] = 15.30

    ws["A7"] = "Coupling"
    ws["B7"] = 95
    ws["C7"] = 8.08

    ws["A8"] = "Bearing"
    ws["B8"] = 320
    ws["C8"] = 27.20

    ws["A9"] = "Washer"
    ws["B9"] = 12
    ws["C9"] = 1.02

    ws["A10"] = "Bolt Set"
    ws["B10"] = 55
    ws["C10"] = 4.68

    ws["A11"] = "Gasket"
    ws["B11"] = 28
    ws["C11"] = 2.38

    # Adjust column widths for readability
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
