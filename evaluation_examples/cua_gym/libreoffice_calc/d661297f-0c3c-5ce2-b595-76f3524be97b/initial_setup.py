"""
Initial Setup: Score Analysis spreadsheet for conditional formatting task
Task ID: calc_fmt_condfmt_above_average_076
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_condfmt_above_average_076'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Score Analysis ---
    ws = wb.active
    ws.title = 'Score Analysis'

    # Headers
    ws['A1'] = 'Participant'
    ws['B1'] = 'Score'

    # Data: 29 participants with realistic scores from 45-98
    participants = [
        ('Alice Martin', 87),
        ('Brian Thompson', 62),
        ('Catherine Wu', 95),
        ('David Patel', 78),
        ('Elena Rodriguez', 53),
        ('Frank Nguyen', 91),
        ('Grace Kim', 70),
        ('Henry Okafor', 48),
        ('Isabella Chen', 83),
        ('James Rivera', 67),
        ('Karen Johansson', 98),
        ('Liam Kowalski', 55),
        ('Mia Fernandez', 76),
        ('Noah Bakker', 45),
        ('Olivia Singh', 89),
        ('Patrick Muller', 72),
        ('Quinn Larsson', 60),
        ('Rachel Dubois', 93),
        ('Samuel Andersen', 81),
        ('Tessa Hoffmann', 57),
        ('Ursula Petrov', 74),
        ('Victor Nakamura', 88),
        ('Wendy Olsen', 64),
        ('Xander Rossi', 96),
        ('Yuki Tanaka', 51),
        ('Zara Ahmed', 79),
        ('Ethan Brooks', 85),
        ('Fatima Hassan', 66),
        ('George Mitchell', 92),
    ]

    for row_idx, (name, score) in enumerate(participants, start=2):
        ws.cell(row=row_idx, column=1, value=name)
        ws.cell(row=row_idx, column=2, value=score)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
