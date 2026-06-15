"""
Reward Script: Import CSV website traffic data, clean, create weekly summary,
               calculate growth %, add trend line chart, format with alternating colors.
Task ID: calc_wf_023
Domain: libreoffice_calc
Scoring:
  Component 1: File exists with both required sheets (0.15)
  Component 2: CSV data imported and cleaned - no blank Visitors (0.15)
  Component 3: Weekly Summary has correct structure - 12 weeks with columns (0.20)
  Component 4: Growth percentages computed correctly (0.15)
  Component 5: Line chart with trendline on Weekly Summary (0.20)
  Component 6: Alternating row colors on summary table (0.15)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_023'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: File has both required sheets (0.15 points)
    try:
        sheet_names = wb.sheetnames
        has_traffic = any('traffic' in s.lower() for s in sheet_names)
        has_summary = any('summary' in s.lower() or 'weekly' in s.lower() for s in sheet_names)
        if has_traffic and has_summary and len(sheet_names) >= 2:
            print(f"PASS: Component 1 -- Both required sheets found: {sheet_names} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 -- Expected sheets with 'traffic' and 'summary/weekly', found: {sheet_names}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Identify the sheets by name (flexible matching)
    traffic_sheet = None
    summary_sheet = None
    for s in wb.sheetnames:
        if 'traffic' in s.lower():
            traffic_sheet = s
        if 'summary' in s.lower() or 'weekly' in s.lower():
            summary_sheet = s

    if not traffic_sheet or not summary_sheet:
        print(f"CRITICAL: Cannot identify required sheets. Stopping.")
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    ws_traffic = wb[traffic_sheet]
    ws_summary = wb[summary_sheet]

    # Component 2: CSV data imported and cleaned -- no blank Visitors (0.15 points)
    try:
        # Check that traffic sheet has substantial data (original CSV had 90 rows, ~5-8 removed)
        traffic_rows = ws_traffic.max_row - 1  # exclude header
        # Check no blank Visitors in column C (Visitors)
        blank_visitors_count = 0
        for r in range(2, ws_traffic.max_row + 1):
            val = ws_traffic.cell(r, 3).value
            if val is None or str(val).strip() == '':
                blank_visitors_count += 1

        if traffic_rows >= 75 and blank_visitors_count == 0:
            print(f"PASS: Component 2 -- Traffic data imported ({traffic_rows} rows) with no blank Visitors (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 -- Traffic rows: {traffic_rows} (need >=75), blank Visitors: {blank_visitors_count} (need 0)")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Weekly Summary has correct structure (0.20 points)
    try:
        summary_rows = ws_summary.max_row - 1  # exclude header
        summary_cols = ws_summary.max_column

        # Check header has expected columns (flexible)
        headers = [str(ws_summary.cell(1, c).value or '').lower() for c in range(1, summary_cols + 1)]
        has_week_col = any('week' in h for h in headers)
        has_visitors_col = any('visitor' in h or 'total' in h for h in headers)
        has_growth_col = any('growth' in h or '%' in h for h in headers)

        # Check we have approximately 12 weeks of data
        correct_structure = (
            summary_rows >= 10 and summary_rows <= 15 and
            has_week_col and has_visitors_col and has_growth_col
        )

        if correct_structure:
            # Verify visitor totals are numeric and reasonable
            visitors_col_idx = None
            for c in range(1, summary_cols + 1):
                h = str(ws_summary.cell(1, c).value or '').lower()
                if 'visitor' in h or 'total' in h:
                    visitors_col_idx = c
                    break

            numeric_count = 0
            if visitors_col_idx:
                for r in range(2, ws_summary.max_row + 1):
                    val = ws_summary.cell(r, visitors_col_idx).value
                    if isinstance(val, (int, float)) and val > 0:
                        numeric_count += 1

            if numeric_count >= 10:
                print(f"PASS: Component 3 -- Weekly summary has {summary_rows} weeks, correct headers, {numeric_count} numeric visitor totals (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 3 -- Only {numeric_count} numeric visitor values found (need >=10)")
        else:
            print(f"FAIL: Component 3 -- Structure issue: rows={summary_rows}, week_col={has_week_col}, visitors_col={has_visitors_col}, growth_col={has_growth_col}")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Growth percentages computed correctly (0.15 points)
    try:
        growth_col_idx = None
        for c in range(1, ws_summary.max_column + 1):
            h = str(ws_summary.cell(1, c).value or '').lower()
            if 'growth' in h or '%' in h:
                growth_col_idx = c
                break

        if growth_col_idx:
            growth_values = []
            for r in range(2, ws_summary.max_row + 1):
                val = ws_summary.cell(r, growth_col_idx).value
                if isinstance(val, (int, float)):
                    growth_values.append(val)

            # First week should be N/A or missing; remaining should be numeric
            # We expect at least 10 numeric growth values (weeks 3-13)
            if len(growth_values) >= 10:
                # Check that growth values are reasonable percentages (between -1.0 and 2.0)
                reasonable = all(-1.0 <= v <= 2.0 for v in growth_values)
                if reasonable:
                    print(f"PASS: Component 4 -- {len(growth_values)} growth percentages found, all in reasonable range (0.15 pts)")
                    total_score += 0.15
                else:
                    print(f"FAIL: Component 4 -- Growth values out of reasonable range: {growth_values}")
            else:
                print(f"FAIL: Component 4 -- Only {len(growth_values)} numeric growth values (need >=10)")
        else:
            print(f"FAIL: Component 4 -- No growth/percentage column found in headers")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # Component 5: Line chart with trendline on Weekly Summary (0.20 points)
    try:
        charts = ws_summary._charts
        if len(charts) >= 1:
            # Count line charts and trendlines via actual API inspection
            line_chart_count = sum(1 for c in charts if 'Line' in type(c).__name__)
            trendline_count = sum(
                1 for c in charts if 'Line' in type(c).__name__
                for s in c.series
                if hasattr(s, 'trendline') and s.trendline is not None
            )

            if line_chart_count >= 1 and trendline_count >= 1:
                print(f"PASS: Component 5 -- Line chart with trendline found (0.20 pts)")
                total_score += 0.20
            elif line_chart_count >= 1:
                # Partial: line chart exists but no trendline
                print(f"PARTIAL: Component 5 -- Line chart found but no trendline (0.10 pts)")
                total_score += 0.10
            else:
                # Chart exists but not a line chart
                print(f"FAIL: Component 5 -- Chart found but not a LineChart (type: {[type(c).__name__ for c in charts]})")
        else:
            print(f"FAIL: Component 5 -- No charts found on summary sheet")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    # Component 6: Alternating row colors on summary table (0.15 points)
    try:
        # Check that data rows have alternating fill colors
        # Even rows should have one color, odd rows another (or vice versa)
        even_fills = set()
        odd_fills = set()
        for r in range(2, ws_summary.max_row + 1):
            cell = ws_summary.cell(r, 1)
            try:
                fill_rgb = cell.fill.fgColor.rgb if cell.fill.fgColor else None
            except Exception:
                fill_rgb = None

            if fill_rgb and fill_rgb != '00000000':
                if r % 2 == 0:
                    even_fills.add(fill_rgb)
                else:
                    odd_fills.add(fill_rgb)

        # For alternating colors, even rows should have a consistent color
        # and odd rows should have a different consistent color
        has_alternating = (
            len(even_fills) == 1 and
            len(odd_fills) == 1 and
            even_fills != odd_fills
        )

        if has_alternating:
            print(f"PASS: Component 6 -- Alternating row colors detected: even={even_fills}, odd={odd_fills} (0.15 pts)")
            total_score += 0.15
        else:
            # Check via conditional formatting as alternative
            cf_mod_row_count = sum(
                1 for cf in ws_summary.conditional_formatting
                for rule in cf.rules
                if rule.formula and any('MOD' in str(f).upper() and 'ROW' in str(f).upper() for f in rule.formula)
            )

            if cf_mod_row_count >= 1:
                print(f"PASS: Component 6 -- Alternating row formatting via conditional formatting with MOD/ROW (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 6 -- No alternating row colors detected. Even fills: {even_fills}, Odd fills: {odd_fills}")
    except Exception as e:
        print(f"ERROR: Component 6 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
