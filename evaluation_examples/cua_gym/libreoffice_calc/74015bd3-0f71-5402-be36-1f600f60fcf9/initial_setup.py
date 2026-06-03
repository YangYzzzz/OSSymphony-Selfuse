"""
Initial Setup: HR Training Completion Tracker
Task ID: calc_hr_training_completion_012
Domain: libreoffice_calc
Description: Creates a spreadsheet with employee training data.
             Columns F (Completion %) and G (Status) are left empty
             for the agent to fill in.
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_training_completion_012'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Training'

    # --- Row 1: Headers ---
    headers = ['Emp ID', 'Name', 'Department', 'Courses Required', 'Courses Completed', 'Completion %', 'Status']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # --- Rows 2-89: Employee training data (88 employees) ---
    # Realistic employee data across various departments
    departments = [
        'Engineering', 'Marketing', 'Sales', 'HR', 'Finance',
        'Operations', 'Legal', 'IT Support', 'Customer Success', 'Product'
    ]

    employee_data = [
        # (emp_id, name, department, courses_required, courses_completed)
        ('E001', 'Sarah Chen', 'Engineering', 8, 8),
        ('E002', 'Marcus Johnson', 'Marketing', 6, 5),
        ('E003', 'Priya Patel', 'HR', 5, 5),
        ('E004', 'David Kim', 'Finance', 7, 7),
        ('E005', 'Laura Gonzalez', 'Sales', 6, 4),
        ('E006', 'James Okonkwo', 'Operations', 5, 3),
        ('E007', 'Mei-Lin Wu', 'Engineering', 8, 8),
        ('E008', 'Aisha Thompson', 'Legal', 9, 9),
        ('E009', 'Robert Vasquez', 'IT Support', 6, 6),
        ('E010', 'Hannah Fischer', 'Customer Success', 5, 2),
        ('E011', 'Oliver Bennett', 'Product', 7, 7),
        ('E012', 'Fatima Al-Rashid', 'Marketing', 6, 6),
        ('E013', 'Cody Harrison', 'Engineering', 8, 5),
        ('E014', 'Yuki Nakamura', 'HR', 5, 5),
        ('E015', 'Andre Dubois', 'Finance', 7, 4),
        ('E016', 'Samantha Torres', 'Sales', 6, 6),
        ('E017', 'Nathan Burke', 'Operations', 5, 5),
        ('E018', 'Ingrid Sorensen', 'Legal', 9, 6),
        ('E019', 'Rajan Mehta', 'IT Support', 6, 3),
        ('E020', 'Tanya Williams', 'Customer Success', 5, 5),
        ('E021', 'Liam O\'Brien', 'Product', 7, 7),
        ('E022', 'Zara Ahmed', 'Engineering', 8, 8),
        ('E023', 'Brett Coleman', 'Marketing', 6, 2),
        ('E024', 'Nadia Kowalski', 'Finance', 7, 7),
        ('E025', 'Ethan Clarke', 'Sales', 6, 6),
        ('E026', 'Grace Liu', 'HR', 5, 4),
        ('E027', 'Tyler Nguyen', 'Operations', 5, 5),
        ('E028', 'Alicia Romero', 'Legal', 9, 9),
        ('E029', 'Kevin Park', 'IT Support', 6, 6),
        ('E030', 'Diana Petit', 'Customer Success', 5, 1),
        ('E031', 'Carlos Moreno', 'Engineering', 8, 6),
        ('E032', 'Sophie Martin', 'Product', 7, 5),
        ('E033', 'Jason Reed', 'Marketing', 6, 6),
        ('E034', 'Amara Osei', 'Finance', 7, 7),
        ('E035', 'Patrick Walsh', 'Sales', 6, 3),
        ('E036', 'Elena Volkov', 'HR', 5, 5),
        ('E037', 'Marcus Lee', 'Operations', 5, 4),
        ('E038', 'Chiara Rossi', 'Legal', 9, 9),
        ('E039', 'Abdullah Hassan', 'IT Support', 6, 6),
        ('E040', 'Rebecca Stone', 'Customer Success', 5, 5),
        ('E041', 'Felix Wagner', 'Engineering', 8, 8),
        ('E042', 'Isabelle Tremblay', 'Product', 7, 7),
        ('E043', 'Kwame Asante', 'Marketing', 6, 4),
        ('E044', 'Hana Yoshida', 'Finance', 7, 7),
        ('E045', 'Logan Campbell', 'Sales', 6, 6),
        ('E046', 'Maria Santos', 'HR', 5, 2),
        ('E047', 'Vincent Leroy', 'Operations', 5, 5),
        ('E048', 'Sasha Petrov', 'Legal', 9, 7),
        ('E049', 'Jasmine Howard', 'IT Support', 6, 6),
        ('E050', 'William Scott', 'Customer Success', 5, 5),
        ('E051', 'Nina Bergstrom', 'Engineering', 8, 4),
        ('E052', 'Omar Abdullah', 'Product', 7, 7),
        ('E053', 'Alicia Pearce', 'Marketing', 6, 6),
        ('E054', 'Thomas Mueller', 'Finance', 7, 6),
        ('E055', 'Vanessa Diaz', 'Sales', 6, 6),
        ('E056', 'Henry Wilson', 'HR', 5, 5),
        ('E057', 'Lily Zhang', 'Operations', 5, 5),
        ('E058', 'Bruce Lawson', 'Legal', 9, 9),
        ('E059', 'Serena Mbeki', 'IT Support', 6, 3),
        ('E060', 'Aaron Phillips', 'Customer Success', 5, 5),
        ('E061', 'Clara Jensen', 'Engineering', 8, 8),
        ('E062', 'Ricardo Alves', 'Product', 7, 4),
        ('E063', 'Nour El-Amin', 'Marketing', 6, 6),
        ('E064', 'Sandra Hoffman', 'Finance', 7, 7),
        ('E065', 'Derek Mwangi', 'Sales', 6, 2),
        ('E066', 'Petra Kovacs', 'HR', 5, 5),
        ('E067', 'Samuel Green', 'Operations', 5, 5),
        ('E068', 'Yolanda Cruz', 'Legal', 9, 8),
        ('E069', 'Tobias Schreiber', 'IT Support', 6, 6),
        ('E070', 'Mei Tanaka', 'Customer Success', 5, 5),
        ('E071', 'Brendan Fox', 'Engineering', 8, 8),
        ('E072', 'Ananya Sharma', 'Product', 7, 7),
        ('E073', 'Michel Dupont', 'Marketing', 6, 6),
        ('E074', 'Erin McCarthy', 'Finance', 7, 3),
        ('E075', 'Ibrahim Al-Farsi', 'Sales', 6, 6),
        ('E076', 'Julia Becker', 'HR', 5, 5),
        ('E077', 'Dante Ricci', 'Operations', 5, 1),
        ('E078', 'Chloe Armstrong', 'Legal', 9, 9),
        ('E079', 'Stefan Andersen', 'IT Support', 6, 6),
        ('E080', 'Kenji Matsumoto', 'Customer Success', 5, 4),
        ('E081', 'Rachel Foster', 'Engineering', 8, 8),
        ('E082', 'Alejandro Reyes', 'Product', 7, 6),
        ('E083', 'Fatou Diallo', 'Marketing', 6, 6),
        ('E084', 'Geoffrey Hart', 'Finance', 7, 7),
        ('E085', 'Catalina Ruiz', 'Sales', 6, 5),
        ('E086', 'Nikolai Ivanov', 'HR', 5, 5),
        ('E087', 'Abby Sullivan', 'Operations', 5, 5),
        ('E088', 'Leo Fontaine', 'Legal', 9, 0),
    ]

    for row_idx, (emp_id, name, dept, req, comp) in enumerate(employee_data, 2):
        ws.cell(row=row_idx, column=1, value=emp_id)
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=dept)
        ws.cell(row=row_idx, column=4, value=req)
        ws.cell(row=row_idx, column=5, value=comp)
        # Columns F (6) and G (7) intentionally left empty

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Training')
    print(f'  Rows: 1 header + 88 data rows (rows 2-89)')
    print(f'  Columns F and G are empty (to be filled by agent)')


create_initial()
