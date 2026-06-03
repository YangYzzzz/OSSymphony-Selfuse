"""
Reward Script: HR Training Completion Tracker Setup
Task ID: calc_hr_training_completion_012
Domain: libreoffice_calc
Scoring:
  Component 1: F2:F89 have IFERROR(E/D,0) completion percentage formulas  (0.30 pts)
  Component 2: F2:F89 have percentage number format '0%'                   (0.10 pts)
  Component 3: G2:G89 have IF(F=1,"Complete","Incomplete") status formulas (0.25 pts)
  Component 4: Data rows A2:G89 sorted by completion % (E/D ratio) desc   (0.20 pts)
  Component 5: Conditional formatting on G2:G89 with correct colors        (0.15 pts)
               - "Complete"   -> #70AD47 (green)  FF70AD47
               - "Incomplete" -> #FFC000 (amber)  FFFFC000
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_hr_training_completion_012'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Training' sheet must exist
    if 'Training' not in wb.sheetnames:
        print("CRITICAL: 'Training' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Training']

    # -------------------------------------------------------------------------
    # Component 1: F2:F89 contain IFERROR(E/D,0) completion % formulas (0.30 pts)
    # Fails on initial (all None), passes on golden (all have the formula)
    # -------------------------------------------------------------------------
    try:
        formula_count = 0
        total_cells = 88  # rows 2-89 = 88 cells
        formula_errors = []

        for r in range(2, 90):
            val = ws.cell(row=r, column=6).value  # Column F
            if val is None:
                formula_errors.append(f"F{r} is empty")
                continue
            if isinstance(val, str):
                # Normalize formula for comparison: remove spaces, uppercase
                norm = val.upper().replace(" ", "")
                # Accept =IFERROR(En/Dn,0) pattern
                expected = f"=IFERROR(E{r}/D{r},0)"
                expected_norm = expected.upper().replace(" ", "")
                if norm == expected_norm:
                    formula_count += 1
                else:
                    formula_errors.append(f"F{r}: unexpected formula {repr(val)}")
            else:
                formula_errors.append(f"F{r}: not a formula string, got {repr(val)}")

        # Require all 88 cells to have correct formula
        if formula_count == total_cells:
            print(f"PASS: Component 1 — All F2:F89 have =IFERROR(En/Dn,0) formula ({total_cells}/{total_cells}) (0.30 pts)")
            total_score += 0.30
        elif formula_count >= total_cells * 0.9:
            print(f"PARTIAL: Component 1 — {formula_count}/{total_cells} F cells have correct formula — first errors: {formula_errors[:3]}")
            # Partial credit: 0.15 if 90%+ correct
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — Only {formula_count}/{total_cells} F cells have correct formula — errors: {formula_errors[:5]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: F2:F89 have percentage number format '0%' (0.10 pts)
    # Fails on initial (General format), passes on golden (0% format)
    # -------------------------------------------------------------------------
    try:
        pct_format_count = 0
        for r in range(2, 90):
            fmt = ws.cell(row=r, column=6).number_format
            # Accept '0%' or '0.0%' or '0.00%' as valid percentage formats
            if fmt and '%' in fmt:
                pct_format_count += 1

        if pct_format_count == 88:
            print(f"PASS: Component 2 — All F2:F89 have percentage number format (0.10 pts)")
            total_score += 0.10
        elif pct_format_count >= 70:
            print(f"PARTIAL: Component 2 — {pct_format_count}/88 F cells have percentage format")
            total_score += 0.05
        else:
            print(f"FAIL: Component 2 — Only {pct_format_count}/88 F cells have percentage format (expected all 88)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: G2:G89 contain IF(Fn=1,"Complete","Incomplete") formulas (0.25 pts)
    # Fails on initial (all None), passes on golden
    # -------------------------------------------------------------------------
    try:
        g_formula_count = 0
        g_errors = []

        for r in range(2, 90):
            val = ws.cell(row=r, column=7).value  # Column G
            if val is None:
                g_errors.append(f"G{r} is empty")
                continue
            if isinstance(val, str):
                norm = val.upper().replace(" ", "").replace('"', '"').replace('"', '"')
                expected = f'=IF(F{r}=1,"Complete","Incomplete")'
                expected_norm = expected.upper().replace(" ", "")
                if norm == expected_norm:
                    g_formula_count += 1
                else:
                    g_errors.append(f"G{r}: unexpected formula {repr(val)}")
            else:
                g_errors.append(f"G{r}: not a formula, got {repr(val)}")

        if g_formula_count == 88:
            print(f"PASS: Component 3 — All G2:G89 have =IF(Fn=1,\"Complete\",\"Incomplete\") formula (0.25 pts)")
            total_score += 0.25
        elif g_formula_count >= 88 * 0.9:
            print(f"PARTIAL: Component 3 — {g_formula_count}/88 G cells have correct formula — errors: {g_errors[:3]}")
            total_score += 0.12
        else:
            print(f"FAIL: Component 3 — Only {g_formula_count}/88 G cells have correct formula — errors: {g_errors[:5]}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: Data rows A2:G89 sorted by completion % (E/D ratio) descending (0.20 pts)
    # Fails on initial (original order E001,E002,...), passes on golden (sorted)
    # Strategy: compute E/D ratio for each row and verify descending order
    # -------------------------------------------------------------------------
    try:
        ratios = []
        compute_errors = []
        for r in range(2, 90):
            d_val = ws.cell(row=r, column=4).value  # Courses Required
            e_val = ws.cell(row=r, column=5).value  # Courses Completed
            if d_val is not None and e_val is not None and isinstance(d_val, (int, float)) and isinstance(e_val, (int, float)) and d_val > 0:
                ratios.append((r, float(e_val) / float(d_val)))
            else:
                compute_errors.append(f"Row {r}: D={d_val}, E={e_val}")

        if compute_errors:
            print(f"WARNING: Component 4 — Could not compute ratio for {len(compute_errors)} rows: {compute_errors[:3]}")

        # Check if ratios are in non-increasing (descending) order
        if len(ratios) >= 2:
            out_of_order = 0
            for i in range(len(ratios) - 1):
                # Allow small floating-point tolerance
                if ratios[i][1] < ratios[i+1][1] - 1e-9:
                    out_of_order += 1

            if out_of_order == 0:
                print(f"PASS: Component 4 — Data rows are sorted by completion % descending "
                      f"(top={ratios[0][1]:.4f}, bottom={ratios[-1][1]:.4f}) (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 4 — {out_of_order} order violations found in completion % sort "
                      f"(e.g. row {ratios[0][0]}={ratios[0][1]:.4f}, ...; not descending)")
        else:
            print(f"FAIL: Component 4 — Could not verify sort order (insufficient ratio data)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -------------------------------------------------------------------------
    # Component 5: Conditional formatting on G2:G89 with correct colors (0.15 pts)
    # Complete -> FF70AD47 (green), Incomplete -> FFFFC000 (amber)
    # Fails on initial (no CF), passes on golden
    # -------------------------------------------------------------------------
    try:
        cf_list = ws.conditional_formatting
        cf_rules_dict = cf_list._cf_rules

        # Find CF for G2:G89
        target_range_key = None
        for key in cf_rules_dict.keys():
            key_str = str(key).upper().replace(" ", "")
            if "G2:G89" in key_str or "G2" in key_str:
                target_range_key = key
                break

        if target_range_key is None:
            print("FAIL: Component 5 — No conditional formatting found on column G")
        else:
            rules = cf_rules_dict[target_range_key]
            complete_color_ok = False
            incomplete_color_ok = False

            for rule in rules:
                # Check formula and fill color
                formula_str = ""
                if hasattr(rule, 'formula') and rule.formula:
                    formula_str = " ".join(str(f) for f in rule.formula).upper().replace(" ", "").replace('"', '').replace("'", "")

                fill_rgb = None
                if hasattr(rule, 'dxf') and rule.dxf is not None:
                    try:
                        if hasattr(rule.dxf, 'fill') and rule.dxf.fill is not None:
                            fg = rule.dxf.fill.fgColor
                            if fg.type == 'rgb':
                                fill_rgb = fg.rgb
                    except Exception:
                        pass

                # Complete: formula contains "Complete" but NOT "Incomplete", color FF70AD47 (green)
                # Note: "INCOMPLETE" contains "COMPLETE", so check for exact "INCOMPLETE" first
                if "INCOMPLETE" in formula_str and fill_rgb:
                    if fill_rgb.upper() == "FFFFC000":
                        incomplete_color_ok = True
                    else:
                        print(f"FAIL: Component 5 — 'Incomplete' rule has wrong color: {fill_rgb} (expected FFFFC000)")
                elif "COMPLETE" in formula_str and fill_rgb:
                    # Only "Complete" (not "Incomplete") rule
                    if fill_rgb.upper() == "FF70AD47":
                        complete_color_ok = True
                    else:
                        print(f"FAIL: Component 5 — 'Complete' rule has wrong color: {fill_rgb} (expected FF70AD47)")

            if complete_color_ok and incomplete_color_ok:
                print(f"PASS: Component 5 — Conditional formatting on G2:G89: "
                      f"Complete=FF70AD47 (green), Incomplete=FFFFC000 (amber) (0.15 pts)")
                total_score += 0.15
            elif complete_color_ok or incomplete_color_ok:
                print(f"PARTIAL: Component 5 — Only one CF rule correct: "
                      f"complete_ok={complete_color_ok}, incomplete_ok={incomplete_color_ok}")
                total_score += 0.07
            else:
                print(f"FAIL: Component 5 — CF rules found but colors incorrect: "
                      f"complete_ok={complete_color_ok}, incomplete_ok={incomplete_color_ok}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
