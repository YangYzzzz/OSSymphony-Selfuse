"""
Initial Setup: Add absolute and percentage change columns to city population growth table
Task ID: osworld_calc_annual_pct_change_012
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_annual_pct_change_012'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Population Data ---
    ws = wb.active
    ws.title = "Population Data"

    # Headers: only City, Population 2020, Population 2025
    # MUST NOT include Absolute Change or Percentage Change columns
    headers = ['City', 'Population 2020', 'Population 2025']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic city population data (12 cities)
    data = [
        ['Shanghai',    24870895, 28516904],
        ['Beijing',     21893095, 24900000],
        ['Tokyo',       13960000, 13515271],
        ['Delhi',       16787941, 20667656],
        ['Mumbai',      12442373, 13085000],
        ['São Paulo',   12325232, 12325232],
        ['Mexico City', 9209944,  9867000],
        ['Cairo',       10107125, 11190000],
        ['Dhaka',       8906039,  10278882],
        ['Osaka',       2691185,  2756000],
        ['Karachi',     14910352, 16093786],
        ['Lagos',       13463293, 15388000],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
