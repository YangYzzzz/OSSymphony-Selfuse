"""
Reward Script: Add absolute and percentage change columns to population growth table
Task ID: osworld_calc_annual_pct_change_012
Domain: libreoffice_calc
Scoring:
  - Component 1: Column D header is present (e.g., 'Absolute Change') — 0.1 pts
  - Component 2: Column E header is present (e.g., 'Percentage Change') — 0.1 pts
  - Component 3: All 12 data rows in column D have absolute change formula (=C#-B#) — 0.4 pts
  - Component 4: All 12 data rows in column E have percentage change formula (=(C#-B#)/B#*100) — 0.4 pts
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_annual_pct_change_012'

# Expected data rows: rows 2 through 13 (12 cities)
DATA_ROWS = list(range(2, 14))


def normalize_formula(formula_str):
    """Normalize formula string for comparison: uppercase, remove spaces."""
    if formula_str is None:
        return ''
    return str(formula_str).upper().replace(' ', '')


def check_abs_formula(cell_value, row):
    """
    Check that column D cell contains an absolute change formula.
    Expected pattern: =C{row}-B{row}  (case-insensitive, spaces ignored)
    """
    if cell_value is None:
        return False
    norm = normalize_formula(cell_value)
    # Must start with '=' and contain C{row}-B{row}
    expected = f'=C{row}-B{row}'
    return norm == expected.upper()


def check_pct_formula(cell_value, row):
    """
    Check that column E cell contains a percentage change formula.
    Expected pattern: =(C{row}-B{row})/B{row}*100  (case-insensitive, spaces ignored)
    Also accept alternative equivalent formulas like =(C{row}-B{row})/B{row}*100
    """
    if cell_value is None:
        return False
    norm = normalize_formula(cell_value)
    expected = f'=(C{row}-B{row})/B{row}*100'
    return norm == expected.upper()


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Precondition: verify file has at least 5 columns (task changes from 3 to 5 columns)
    if ws.max_column < 4:
        print("FAIL: File has fewer than 4 columns — columns D and E not added yet")
        print("\nScore: 0.0/1.0")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Column D header is present (0.1 points)
    # Task requires a header for the absolute change column in D1
    try:
        d1_value = ws.cell(row=1, column=4).value
        if d1_value is not None and len(str(d1_value).strip()) > 0:
            # Accept any non-empty header containing relevant keywords or any header at all
            d1_str = str(d1_value).strip().lower()
            # Check it's a meaningful header (not a formula)
            if not d1_str.startswith('='):
                print(f"PASS: Component 1 — Column D header present: '{d1_value}' (0.1 pts)")
                total_score += 0.1
            else:
                print(f"FAIL: Component 1 — Column D header appears to be formula, not label: '{d1_value}'")
        else:
            print(f"FAIL: Component 1 — Column D header missing (D1 is empty)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Column E header is present (0.1 points)
    # Task requires a header for the percentage change column in E1
    try:
        e1_value = ws.cell(row=1, column=5).value if ws.max_column >= 5 else None
        if e1_value is not None and len(str(e1_value).strip()) > 0:
            e1_str = str(e1_value).strip().lower()
            if not e1_str.startswith('='):
                print(f"PASS: Component 2 — Column E header present: '{e1_value}' (0.1 pts)")
                total_score += 0.1
            else:
                print(f"FAIL: Component 2 — Column E header appears to be formula, not label: '{e1_value}'")
        else:
            print(f"FAIL: Component 2 — Column E header missing (E1 is empty)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: All 12 data rows in column D have absolute change formula =C#-B# (0.4 points)
    try:
        abs_correct = 0
        abs_total = len(DATA_ROWS)
        for row in DATA_ROWS:
            cell_val = ws.cell(row=row, column=4).value
            if check_abs_formula(cell_val, row):
                abs_correct += 1
            else:
                print(f"  FAIL row {row} col D: expected '=C{row}-B{row}', found: {repr(cell_val)}")

        if abs_correct == abs_total:
            print(f"PASS: Component 3 — All {abs_total} absolute change formulas correct in col D (0.4 pts)")
            total_score += 0.4
        elif abs_correct > 0:
            partial = round(0.4 * abs_correct / abs_total, 3)
            print(f"PARTIAL: Component 3 — {abs_correct}/{abs_total} absolute change formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No correct absolute change formulas found in column D")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: All 12 data rows in column E have percentage change formula =(C#-B#)/B#*100 (0.4 points)
    try:
        pct_correct = 0
        pct_total = len(DATA_ROWS)
        for row in DATA_ROWS:
            cell_val = ws.cell(row=row, column=5).value if ws.max_column >= 5 else None
            if check_pct_formula(cell_val, row):
                pct_correct += 1
            else:
                print(f"  FAIL row {row} col E: expected '=(C{row}-B{row})/B{row}*100', found: {repr(cell_val)}")

        if pct_correct == pct_total:
            print(f"PASS: Component 4 — All {pct_total} percentage change formulas correct in col E (0.4 pts)")
            total_score += 0.4
        elif pct_correct > 0:
            partial = round(0.4 * pct_correct / pct_total, 3)
            print(f"PARTIAL: Component 4 — {pct_correct}/{pct_total} percentage change formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — No correct percentage change formulas found in column E")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the given env
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
