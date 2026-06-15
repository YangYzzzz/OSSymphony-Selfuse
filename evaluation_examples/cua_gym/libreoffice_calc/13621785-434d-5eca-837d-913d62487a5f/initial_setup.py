"""
Initial Setup: Apply border scheme to Master Template sheet
Task ID: calc_ggf_034
Domain: libreoffice_calc

Creates a 100-row by 10-column employee data table on a 'Master Template' sheet
with NO borders applied. The agent's task is to apply the specified border scheme.
"""

import os
import shlex
import subprocess
import time
import random
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_034'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Master Template'

    # --- Headers (Row 1) ---
    headers = [
        'Employee ID', 'Full Name', 'Department', 'Position',
        'Annual Salary', 'Start Date', 'Performance Score',
        'Region', 'Employment Status', 'Notes'
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # --- Realistic employee data (rows 2-100) ---
    first_names = [
        'Sarah', 'Marcus', 'Elena', 'David', 'Priya', 'James', 'Aiko',
        'Robert', 'Fatima', 'Carlos', 'Wei', 'Jennifer', 'Oluwaseun',
        'Thomas', 'Maria', 'Ahmed', 'Sofia', 'Michael', 'Yuki', 'Anna',
        'Daniel', 'Lena', 'Patrick', 'Mei', 'Christopher', 'Nadia',
        'Kevin', 'Ingrid', 'Samuel', 'Zara'
    ]
    last_names = [
        'Chen', 'Johnson', 'Kowalski', 'Park', 'Sharma', 'Williams',
        'Tanaka', 'Martinez', 'Al-Hassan', 'Rivera', 'Zhang', 'O\'Brien',
        'Adeyemi', 'Mueller', 'Santos', 'Ibrahim', 'Petrov', 'Anderson',
        'Nakamura', 'Svensson', 'Foster', 'Bergmann', 'Kim', 'Liu',
        'Thompson', 'Novak', 'Reed', 'Johansson', 'Okafor', 'Abbas'
    ]
    departments = [
        'Engineering', 'Marketing', 'Finance', 'Human Resources',
        'Operations', 'Sales', 'Research & Development', 'Legal',
        'Customer Success', 'Product Management'
    ]
    positions = [
        'Senior Engineer', 'Marketing Analyst', 'Financial Controller',
        'HR Specialist', 'Operations Manager', 'Sales Representative',
        'Research Scientist', 'Legal Counsel', 'Account Manager',
        'Product Owner', 'Team Lead', 'Director', 'Associate',
        'Vice President', 'Coordinator', 'Consultant', 'Architect',
        'Strategist', 'Administrator', 'Specialist'
    ]
    regions = [
        'North America', 'Europe', 'Asia Pacific', 'Latin America',
        'Middle East', 'Africa'
    ]
    statuses = ['Active', 'Active', 'Active', 'Active', 'On Leave', 'Probation']
    notes_options = [
        'Exceeds expectations consistently',
        'Completed leadership training',
        'Relocated from London office',
        'Mentor for new hires',
        'Project lead for Q2 initiative',
        'Bilingual - Spanish/English',
        'Remote worker since 2024',
        'Due for annual review',
        'Transferred from Singapore',
        'Patent holder - 3 filed',
        '',
        '',
        '',
    ]

    random.seed(42)  # reproducible data
    for row in range(2, 101):
        emp_id = f'EMP-{1000 + row - 1:05d}'
        name = f'{random.choice(first_names)} {random.choice(last_names)}'
        dept = random.choice(departments)
        pos = random.choice(positions)
        salary = round(random.uniform(48000, 185000), 2)
        year = random.randint(2018, 2025)
        month = random.randint(1, 12)
        day = random.randint(1, 28)
        start_date = f'{year}-{month:02d}-{day:02d}'
        perf = round(random.uniform(2.5, 5.0), 1)
        region = random.choice(regions)
        status = random.choice(statuses)
        note = random.choice(notes_options)

        ws.cell(row=row, column=1, value=emp_id)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=dept)
        ws.cell(row=row, column=4, value=pos)
        ws.cell(row=row, column=5, value=salary)
        ws.cell(row=row, column=6, value=start_date)
        ws.cell(row=row, column=7, value=perf)
        ws.cell(row=row, column=8, value=region)
        ws.cell(row=row, column=9, value=status)
        ws.cell(row=row, column=10, value=note)

    # Set reasonable column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 30

    # NO borders applied - that is the agent's task
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
