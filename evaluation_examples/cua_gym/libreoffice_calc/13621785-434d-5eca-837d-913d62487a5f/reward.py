"""
Reward Script: Apply specific border scheme to Master Template sheet
Task ID: calc_ggf_034
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Thick dark-blue outer border on A1:J100
  Component 2 (0.30): Medium gray horizontal inner borders between rows
  Component 3 (0.20): No vertical inner borders between columns B-I
  Component 4 (0.20): Double bottom border on row 1 (header)
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_034'

# Persistence hook: save any unsaved LibreOffice edits
def persist_app_state():
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify border scheme on Master Template sheet.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that 'Master Template' sheet exists
    if 'Master Template' not in wb.sheetnames:
        print("CRITICAL: 'Master Template' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Master Template']

    # Helper to get border side info
    def get_side(cell, side_name):
        """Returns (style, color_rgb) for a border side."""
        s = getattr(cell.border, side_name)
        style = s.style  # None, 'thin', 'medium', 'thick', 'double', etc.
        color = None
        if s.color:
            color = s.color.rgb  # ARGB string like '00003366'
        return style, color

    def color_matches(color_rgb, expected_6char):
        """Check if an ARGB color string matches expected 6-char RGB (ignoring alpha)."""
        if color_rgb is None:
            return False
        # color_rgb could be 6 or 8 chars
        rgb_part = color_rgb[-6:]  # last 6 chars = RGB
        return rgb_part.upper() == expected_6char.upper()

    DARK_BLUE = '003366'
    GRAY = '808080'

    # =========================================================
    # Component 1: Thick dark-blue outer border (0.30 points)
    # Check: top edge of row 1, bottom edge of row 100,
    #        left edge of column A, right edge of column J
    # =========================================================
    try:
        outer_checks_pass = 0
        outer_checks_total = 0

        # Top edge: cells A1..J1 should have top=thick/dark-blue
        for col in range(1, 11):
            cell = ws.cell(row=1, column=col)
            style, color = get_side(cell, 'top')
            outer_checks_total += 1
            if style == 'thick' and color_matches(color, DARK_BLUE):
                outer_checks_pass += 1

        # Bottom edge: cells A100..J100 should have bottom=thick/dark-blue
        for col in range(1, 11):
            cell = ws.cell(row=100, column=col)
            style, color = get_side(cell, 'bottom')
            outer_checks_total += 1
            if style == 'thick' and color_matches(color, DARK_BLUE):
                outer_checks_pass += 1

        # Left edge: cells A1..A100 should have left=thick/dark-blue
        # Sample rows to avoid checking all 100
        sample_rows = [1, 2, 10, 25, 50, 75, 99, 100]
        for r in sample_rows:
            cell = ws.cell(row=r, column=1)
            style, color = get_side(cell, 'left')
            outer_checks_total += 1
            if style == 'thick' and color_matches(color, DARK_BLUE):
                outer_checks_pass += 1

        # Right edge: cells J1..J100 should have right=thick/dark-blue
        for r in sample_rows:
            cell = ws.cell(row=r, column=10)
            style, color = get_side(cell, 'right')
            outer_checks_total += 1
            if style == 'thick' and color_matches(color, DARK_BLUE):
                outer_checks_pass += 1

        outer_ratio = outer_checks_pass / outer_checks_total if outer_checks_total > 0 else 0
        if outer_ratio >= 0.9:
            print(f"PASS: Component 1 — Thick dark-blue outer border ({outer_checks_pass}/{outer_checks_total} checks) (0.30 pts)")
            total_score += 0.30
        elif outer_ratio >= 0.5:
            partial = round(0.30 * outer_ratio, 2)
            print(f"PARTIAL: Component 1 — Outer border partially correct ({outer_checks_pass}/{outer_checks_total}) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Outer border missing or wrong ({outer_checks_pass}/{outer_checks_total})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================
    # Component 2: Medium gray horizontal inner borders (0.30 points)
    # Check: interior cells (rows 2-99) should have top and bottom
    #        borders that are medium/gray (horizontal separators)
    # =========================================================
    try:
        horiz_checks_pass = 0
        horiz_checks_total = 0

        # Sample interior cells across different rows and columns
        sample_interior_rows = [3, 10, 20, 30, 50, 70, 90, 99]
        sample_interior_cols = [2, 5, 8]  # B, E, H (interior columns)

        for r in sample_interior_rows:
            for c in sample_interior_cols:
                cell = ws.cell(row=r, column=c)
                # Check top border
                style_t, color_t = get_side(cell, 'top')
                horiz_checks_total += 1
                if style_t == 'medium' and color_matches(color_t, GRAY):
                    horiz_checks_pass += 1

                # Check bottom border
                style_b, color_b = get_side(cell, 'bottom')
                horiz_checks_total += 1
                if style_b == 'medium' and color_matches(color_b, GRAY):
                    horiz_checks_pass += 1

        horiz_ratio = horiz_checks_pass / horiz_checks_total if horiz_checks_total > 0 else 0
        if horiz_ratio >= 0.9:
            print(f"PASS: Component 2 — Medium gray horizontal inner borders ({horiz_checks_pass}/{horiz_checks_total}) (0.30 pts)")
            total_score += 0.30
        elif horiz_ratio >= 0.5:
            partial = round(0.30 * horiz_ratio, 2)
            print(f"PARTIAL: Component 2 — Horizontal inner borders partially correct ({horiz_checks_pass}/{horiz_checks_total}) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Horizontal inner borders missing or wrong ({horiz_checks_pass}/{horiz_checks_total})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================
    # Component 3: No vertical inner borders B-I, BUT horizontal
    #              borders ARE present (0.20 points)
    # This is a compound check: horizontal borders must exist (proving
    # borders were applied) AND vertical borders must be absent.
    # This ensures the check FAILS on initial_env (which has no borders at all).
    # =========================================================
    try:
        no_vert_pass = 0
        no_vert_total = 0
        has_horiz = 0
        horiz_checked = 0

        # Sample interior cells (columns B=2 through I=9, various rows)
        sample_rows_vert = [2, 10, 25, 50, 75, 99]
        sample_cols_vert = [2, 3, 5, 7, 9]  # B, C, E, G, I

        for r in sample_rows_vert:
            for c in sample_cols_vert:
                cell = ws.cell(row=r, column=c)
                style_l, _ = get_side(cell, 'left')
                style_r, _ = get_side(cell, 'right')
                no_vert_total += 1
                # Should have no left or right border (style is None)
                if style_l is None and style_r is None:
                    no_vert_pass += 1
                # Also check that horizontal borders exist (proving borders were applied)
                style_t, _ = get_side(cell, 'top')
                style_b, _ = get_side(cell, 'bottom')
                horiz_checked += 1
                if style_t is not None or style_b is not None:
                    has_horiz += 1

        no_vert_ratio = no_vert_pass / no_vert_total if no_vert_total > 0 else 0
        has_horiz_ratio = has_horiz / horiz_checked if horiz_checked > 0 else 0

        # Must have BOTH: no vertical borders AND horizontal borders present
        if no_vert_ratio >= 0.9 and has_horiz_ratio >= 0.8:
            print(f"PASS: Component 3 — No vertical inner borders B-I with horizontal borders present ({no_vert_pass}/{no_vert_total} no-vert, {has_horiz}/{horiz_checked} has-horiz) (0.20 pts)")
            total_score += 0.20
        elif no_vert_ratio >= 0.5 and has_horiz_ratio >= 0.5:
            partial = round(0.20 * min(no_vert_ratio, has_horiz_ratio), 2)
            print(f"PARTIAL: Component 3 — Partially correct ({no_vert_pass}/{no_vert_total} no-vert, {has_horiz}/{horiz_checked} has-horiz) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Either vertical borders found or no horizontal borders ({no_vert_pass}/{no_vert_total} no-vert, {has_horiz}/{horiz_checked} has-horiz)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================
    # Component 4: Double bottom border on row 1 (0.20 points)
    # Check: cells A1..J1 should have bottom=double border
    # =========================================================
    try:
        double_pass = 0
        double_total = 0

        for col in range(1, 11):
            cell = ws.cell(row=1, column=col)
            style_b, color_b = get_side(cell, 'bottom')
            double_total += 1
            if style_b == 'double':
                double_pass += 1

        double_ratio = double_pass / double_total if double_total > 0 else 0
        if double_ratio >= 0.9:
            print(f"PASS: Component 4 — Double bottom border on row 1 ({double_pass}/{double_total}) (0.20 pts)")
            total_score += 0.20
        elif double_ratio >= 0.5:
            partial = round(0.20 * double_ratio, 2)
            print(f"PARTIAL: Component 4 — Double border partially applied ({double_pass}/{double_total}) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Double bottom border on row 1 missing ({double_pass}/{double_total})")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state()

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
