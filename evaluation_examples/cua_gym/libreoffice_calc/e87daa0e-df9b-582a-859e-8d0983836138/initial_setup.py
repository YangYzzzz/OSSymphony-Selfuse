"""
Initial Setup: Set custom background color for Total row
Task ID: calc_gfl_070
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_070'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    # --- Headers (Row 1) ---
    headers = ["Category", "Q1 Revenue", "Q2 Revenue", "Q3 Revenue", "Q4 Revenue", "Annual Total"]
    header_font = Font(name="Arial", size=11, bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # --- Data (Rows 2-24) ---
    data = [
        ["Electronics", 125400, 138200, 142800, 167500],
        ["Clothing & Apparel", 89300, 94100, 87600, 112400],
        ["Home & Garden", 67800, 72500, 81200, 95300],
        ["Sports Equipment", 45200, 52800, 58400, 43900],
        ["Books & Media", 32100, 28700, 31500, 38200],
        ["Automotive Parts", 78600, 82300, 79100, 91400],
        ["Health & Beauty", 54300, 59800, 63200, 71500],
        ["Toys & Games", 41200, 38500, 44700, 68900],
        ["Office Supplies", 23800, 25100, 24600, 27300],
        ["Food & Beverages", 96500, 101200, 98700, 115800],
        ["Pet Supplies", 18700, 21300, 23800, 26100],
        ["Jewelry & Watches", 63400, 58900, 67200, 82500],
        ["Musical Instruments", 15200, 17800, 16500, 19400],
        ["Garden Tools", 28900, 35200, 38700, 22100],
        ["Kitchen Appliances", 52100, 48700, 55300, 61800],
        ["Furniture", 87400, 92100, 88600, 103200],
        ["Outdoor Recreation", 34600, 41200, 47800, 31500],
        ["Baby Products", 29800, 31500, 33200, 36700],
        ["Art Supplies", 12400, 14200, 13800, 16500],
        ["Hardware & Tools", 43700, 47200, 51800, 55300],
        ["Travel Accessories", 19500, 22800, 25400, 28100],
        ["Cleaning Supplies", 16300, 17800, 18500, 20200],
        ["Stationery", 11200, 12500, 13100, 14800],
    ]

    currency_format = '$#,##0'
    data_alignment = Alignment(horizontal="right", vertical="center")
    cat_alignment = Alignment(horizontal="left", vertical="center")

    for r, row_data in enumerate(data, 2):
        # Category column
        cell = ws.cell(row=r, column=1, value=row_data[0])
        cell.alignment = cat_alignment
        cell.border = thin_border

        # Revenue columns (Q1-Q4)
        for c, val in enumerate(row_data[1:], 2):
            cell = ws.cell(row=r, column=c, value=val)
            cell.number_format = currency_format
            cell.alignment = data_alignment
            cell.border = thin_border

        # Annual Total formula (column 6)
        cell = ws.cell(row=r, column=6, value=f'=SUM(B{r}:E{r})')
        cell.number_format = currency_format
        cell.alignment = data_alignment
        cell.border = thin_border

    # --- TOTAL Row (Row 25) ---
    total_font = Font(name="Arial", size=11, bold=True)
    ws.cell(row=25, column=1, value="TOTAL").font = total_font
    ws.cell(row=25, column=1).alignment = cat_alignment
    ws.cell(row=25, column=1).border = thin_border

    for col in range(2, 7):
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(col)
        cell = ws.cell(row=25, column=col, value=f'=SUM({col_letter}2:{col_letter}24)')
        cell.font = total_font
        cell.number_format = currency_format
        cell.alignment = data_alignment
        cell.border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 22
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 15

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
