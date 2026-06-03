"""
Reward Script: Set custom background color #E8F4FD for the 'Total' row (row 25)
Task ID: calc_gfl_070
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): At least 1 cell in A25:F25 has solid fill with color FFE8F4FD
  Component 2 (0.4): All 6 cells in A25:F25 have solid fill with exact color FFE8F4FD
  Component 3 (0.3): Row 25 fully colored AND adjacent rows (1, 24) not colored with FFE8F4FD
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_070'
TARGET_COLOR = 'FFE8F4FD'  # 8-char ARGB: FF (opaque) + E8F4FD


def check_cell_bgcolor(ws, row, col, expected_argb):
    """Check if a cell has a solid fill with the expected ARGB color."""
    try:
        cell = ws.cell(row=row, column=col)
        if cell.fill.patternType != 'solid':
            return False
        return cell.fill.fgColor.rgb == expected_argb
    except Exception:
        return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify 'Summary' sheet exists
    if 'Summary' not in wb.sheetnames:
        print("CRITICAL: 'Summary' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Summary']

    # Count how many cells in row 25, columns 1-6, have the target background color
    colored_count = 0
    for col in range(1, 7):
        if check_cell_bgcolor(ws, 25, col, TARGET_COLOR):
            colored_count += 1

    print(f"INFO: {colored_count}/6 cells in row 25 have background color {TARGET_COLOR}")

    # Component 1: At least 1 cell in A25:F25 has the target color (0.3 points)
    # This checks that the user started applying the color to the Total row.
    try:
        if colored_count >= 1:
            print(f"PASS: Component 1 — At least 1 cell in row 25 has color {TARGET_COLOR} ({colored_count}/6) (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — No cells in row 25 have color {TARGET_COLOR}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All 6 cells in A25:F25 have the exact target color (0.4 points)
    # Full completion: every cell in the Total row is colored.
    try:
        if colored_count == 6:
            print(f"PASS: Component 2 — All 6 cells in row 25 have exact color {TARGET_COLOR} (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 2 — Only {colored_count}/6 cells have color {TARGET_COLOR}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Row 25 fully colored AND adjacent rows not colored with target (0.3 points)
    # Ensures color was applied precisely to row 25, not broadly to multiple rows.
    try:
        if colored_count == 6:
            # Check that row 1 and row 24 do NOT have the target color
            adjacent_colored_count = sum(
                1 for check_row in [1, 24]
                for col in range(1, 7)
                if check_cell_bgcolor(ws, check_row, col, TARGET_COLOR)
            )

            if adjacent_colored_count == 0:
                print(f"PASS: Component 3 — Row 25 fully colored, adjacent rows clean (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 3 — {adjacent_colored_count} cells in adjacent rows also have target color")
        else:
            print(f"FAIL: Component 3 — Row 25 not fully colored ({colored_count}/6), skipping adjacency check")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
