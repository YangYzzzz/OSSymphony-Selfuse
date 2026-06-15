"""
Reward Script: Fix floating-point rounding in invoice tax calculations
Task ID: calc_tbl_044
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.7): All 10 tax cells (D7:D16) use ROUND() wrapper
  - Component 2 (0.3): ROUND formulas have correct structure (2 decimal places, correct cell ref)
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_044'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    The task requires wrapping all tax calculations in D7:D16 with ROUND(x, 2).
    Initial state: =B7*0.0825 (unrounded)
    Golden state: =ROUND(B7*0.0825,2) (properly rounded)
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get the Invoice sheet
    try:
        ws = wb['Invoice']
    except KeyError:
        # Try active sheet as fallback
        ws = wb.active
        if ws is None:
            print("CRITICAL: No accessible sheet found")
            print("REWARD: 0.0")
            return 0.0
        print(f"INFO: 'Invoice' sheet not found, using active sheet: {ws.title}")

    # Tax cell rows to check
    tax_rows = list(range(7, 17))  # D7 through D16

    # ---------------------------------------------------------------
    # Component 1: Tax cells (D7:D16) contain ROUND() wrapper (0.7 pts)
    #   Progressive: 0.07 per cell that has ROUND
    # ---------------------------------------------------------------
    round_count = 0
    try:
        for row in tax_rows:
            cell_val = ws.cell(row=row, column=4).value  # Column D
            if cell_val is not None and isinstance(cell_val, str):
                # Check if the formula uses ROUND (case-insensitive)
                formula_upper = cell_val.upper().replace(" ", "")
                if formula_upper.startswith("=ROUND("):
                    round_count += 1
                    print(f"  PASS: D{row} has ROUND wrapper: {cell_val}")
                else:
                    print(f"  FAIL: D{row} missing ROUND wrapper: {cell_val}")
            else:
                print(f"  FAIL: D{row} is not a formula: {repr(cell_val)}")

        comp1_score = round_count * 0.07
        if round_count == 10:
            print(f"PASS: Component 1 — All 10 tax cells use ROUND() ({comp1_score:.2f} pts)")
        else:
            print(f"PARTIAL: Component 1 — {round_count}/10 tax cells use ROUND() ({comp1_score:.2f} pts)")
        if comp1_score > 0:
            total_score += comp1_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ---------------------------------------------------------------
    # Component 2: ROUND formulas have correct structure (0.3 pts)
    #   Each cell's formula should be =ROUND(B<row>*0.0825,2)
    #   Progressive: 0.03 per correctly structured formula
    # ---------------------------------------------------------------
    correct_structure_count = 0
    try:
        for row in tax_rows:
            cell_val = ws.cell(row=row, column=4).value
            if cell_val is not None and isinstance(cell_val, str):
                formula_clean = cell_val.upper().replace(" ", "")
                # Expected pattern: =ROUND(B<row>*0.0825,2)
                # Allow some flexibility: the multiplication could be in different order
                # and the tax rate might be expressed slightly differently
                expected_pattern = f"=ROUND(B{row}*0.0825,2)"
                alt_pattern = f"=ROUND(0.0825*B{row},2)"
                if (formula_clean == expected_pattern.upper() or
                        formula_clean == alt_pattern.upper()):
                    correct_structure_count += 1
                    print(f"  PASS: D{row} has correct ROUND structure")
                else:
                    # Also accept ROUND with the cell ref and 2 decimal places
                    # even if the inner expression differs slightly
                    round_match = re.match(
                        r'^=ROUND\((.+),\s*2\)$',
                        formula_clean
                    )
                    if round_match:
                        inner = round_match.group(1)
                        # Check inner references B<row> and 0.0825
                        if f"B{row}" in inner and "0.0825" in inner:
                            correct_structure_count += 1
                            print(f"  PASS: D{row} has valid ROUND structure (variant): {cell_val}")
                        else:
                            print(f"  FAIL: D{row} ROUND inner expression incorrect: {cell_val}")
                    else:
                        print(f"  FAIL: D{row} formula structure incorrect: {cell_val}")
            else:
                print(f"  FAIL: D{row} is not a formula string")

        comp2_score = correct_structure_count * 0.03
        if correct_structure_count == 10:
            print(f"PASS: Component 2 — All 10 formulas have correct structure ({comp2_score:.2f} pts)")
        else:
            print(f"PARTIAL: Component 2 — {correct_structure_count}/10 correct structure ({comp2_score:.2f} pts)")
        if comp2_score > 0:
            total_score += comp2_score
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
