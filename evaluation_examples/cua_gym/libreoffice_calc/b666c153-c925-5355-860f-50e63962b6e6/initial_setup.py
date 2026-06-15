"""
Initial Setup: Invoice with floating-point rounding discrepancy in tax calculations
Task ID: calc_tbl_044
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_044'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # ---- Styling definitions ----
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    currency_fmt = '$#,##0.00'
    pct_fmt = '0.00%'
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # ---- Invoice header info ----
    ws.merge_cells("A1:E1")
    ws["A1"] = "TechSupply Co. — Purchase Invoice"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="2F5496")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws["A2"] = "Invoice #:"
    ws["B2"] = "INV-2025-0312"
    ws["A3"] = "Date:"
    ws["B3"] = "2025-03-15"
    ws["A4"] = "Customer:"
    ws["B4"] = "Riverside Engineering Group"
    ws["A2"].font = Font(bold=True)
    ws["A3"].font = Font(bold=True)
    ws["A4"].font = Font(bold=True)

    # ---- Column headers (row 6) ----
    headers = ["Item Description", "Amount", "Tax Rate", "Tax Amount", "Line Total"]
    col_widths = [30, 14, 12, 14, 14]
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=6, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col_idx)].width = w

    # ---- Line items (rows 7-16) ----
    # Items chosen so that 4 have a +0.0025 rounding diff and 6 have 0 diff,
    # producing a net +$0.01 discrepancy in the tax total.
    items = [
        ("Wireless Keyboard", 69.00),
        ("USB-C Docking Station", 189.00),
        ("Ergonomic Mouse Pad", 45.00),
        ("Adjustable Laptop Stand", 109.00),
        ("LED Desk Lamp", 52.00),
        ("Cable Management Kit", 28.00),
        ("Monitor Riser", 76.00),
        ("Webcam HD 1080p", 84.00),
        ("Noise-Canceling Headset", 196.00),
        ("Desk Organizer Set", 76.00),
    ]

    tax_rate = 0.0825  # 8.25%
    data_start = 7
    for i, (desc, amt) in enumerate(items):
        row = data_start + i
        # A: description
        ws.cell(row=row, column=1, value=desc).border = thin_border
        # B: amount
        b_cell = ws.cell(row=row, column=2, value=amt)
        b_cell.number_format = currency_fmt
        b_cell.border = thin_border
        # C: tax rate (display only)
        c_cell = ws.cell(row=row, column=3, value=tax_rate)
        c_cell.number_format = pct_fmt
        c_cell.alignment = Alignment(horizontal="center")
        c_cell.border = thin_border
        # D: tax amount — UNROUNDED formula (this is the bug)
        d_cell = ws.cell(row=row, column=4)
        d_cell.value = f'=B{row}*0.0825'
        d_cell.number_format = currency_fmt
        d_cell.border = thin_border
        # E: line total
        e_cell = ws.cell(row=row, column=5)
        e_cell.value = f'=B{row}+D{row}'
        e_cell.number_format = currency_fmt
        e_cell.border = thin_border

    last_data_row = data_start + len(items) - 1  # row 16

    # ---- Totals row (row 18) ----
    totals_row = last_data_row + 2  # row 18
    ws.cell(row=totals_row, column=1, value="GRAND TOTAL").font = Font(bold=True, size=12)

    # B: subtotal
    b_total = ws.cell(row=totals_row, column=2)
    b_total.value = f'=SUM(B{data_start}:B{last_data_row})'
    b_total.number_format = currency_fmt
    b_total.font = Font(bold=True)
    b_total.border = Border(top=Side(style="double", color="000000"),
                            bottom=Side(style="double", color="000000"))

    # D: tax total (sums unrounded tax values → shows the penny discrepancy)
    d_total = ws.cell(row=totals_row, column=4)
    d_total.value = f'=SUM(D{data_start}:D{last_data_row})'
    d_total.number_format = currency_fmt
    d_total.font = Font(bold=True)
    d_total.border = Border(top=Side(style="double", color="000000"),
                            bottom=Side(style="double", color="000000"))

    # E: grand total
    e_total = ws.cell(row=totals_row, column=5)
    e_total.value = f'=SUM(E{data_start}:E{last_data_row})'
    e_total.number_format = currency_fmt
    e_total.font = Font(bold=True, size=12)
    e_total.border = Border(top=Side(style="double", color="000000"),
                            bottom=Side(style="double", color="000000"))

    # ---- Note about the discrepancy (row 20) ----
    ws.cell(row=20, column=1,
            value="Note: Grand total may show a 1-cent discrepancy due to tax rounding.").font = \
        Font(italic=True, color="FF0000")

    # ---- Save ----
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # ---- GUI startup ----
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
