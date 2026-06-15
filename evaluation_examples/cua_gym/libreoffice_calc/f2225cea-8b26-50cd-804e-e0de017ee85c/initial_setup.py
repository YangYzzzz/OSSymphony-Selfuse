"""
Initial Setup: Fill blank cells in Customer Segment column and create pivot summary
Task ID: osworld_calc_fill_blanks_above_005
Domain: libreoffice_calc

Creates a spreadsheet with order data where column B (Customer Segment) has values
only on the first row of each segment block. The remaining rows in the segment have
blank cells that need to be filled by carrying down.
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_fill_blanks_above_005'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Orders ---
    ws = wb.active
    ws.title = 'Orders'

    # Headers
    headers = ['Order ID', 'Customer Segment', 'Product', 'Amount']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Order data: column B has segment label only on the FIRST row of each segment block.
    # Remaining rows in the same segment block have blank B cells.
    # Segments: Corporate, Consumer, Home Office, Small Business
    data = [
        # Row 2-7: Corporate
        ['ORD-10021', 'Corporate',     'Laptop Pro 15',         3200.00],
        ['ORD-10022', None,            'Wireless Mouse',          29.99],
        ['ORD-10023', None,            'USB-C Docking Station',  189.50],
        ['ORD-10024', None,            'Mechanical Keyboard',     89.00],
        ['ORD-10025', None,            'Monitor 27in 4K',        620.00],
        ['ORD-10026', None,            'Webcam HD 1080p',         74.95],
        # Row 8-13: Consumer
        ['ORD-10027', 'Consumer',      'Bluetooth Headphones',   149.99],
        ['ORD-10028', None,            'Tablet Stand',            34.50],
        ['ORD-10029', None,            'Smart Speaker',           89.00],
        ['ORD-10030', None,            'USB Hub 7-Port',          45.00],
        ['ORD-10031', None,            'Cable Management Kit',    22.75],
        ['ORD-10032', None,            'Phone Charger 65W',       39.99],
        # Row 14-18: Home Office
        ['ORD-10033', 'Home Office',   'Ergonomic Chair',        549.00],
        ['ORD-10034', None,            'Standing Desk Converter', 275.00],
        ['ORD-10035', None,            'Desk Lamp LED',           59.99],
        ['ORD-10036', None,            'Noise Cancelling Earbud', 199.00],
        ['ORD-10037', None,            'Surge Protector 8-Port',  38.50],
        # Row 19-24: Small Business
        ['ORD-10038', 'Small Business','Network Switch 8-Port',  119.00],
        ['ORD-10039', None,            'Label Printer',          139.50],
        ['ORD-10040', None,            'Barcode Scanner',         95.00],
        ['ORD-10041', None,            'Receipt Printer',        249.00],
        ['ORD-10042', None,            'Cash Drawer',             89.00],
        ['ORD-10043', None,            'Card Reader USB',         55.75],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 12

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
